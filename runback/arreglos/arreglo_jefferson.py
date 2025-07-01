#!/usr/bin/env python

import os
import sys
import time

YOUR_PATH = os.path.dirname(os.path.realpath(__file__))
SITE_ROOT = os.path.dirname(os.path.dirname(YOUR_PATH))
SITE_ROOT = os.path.join(SITE_ROOT, '')
your_djangoproject_home = os.path.split(SITE_ROOT)[0]
sys.path.append(your_djangoproject_home)
from django.core.wsgi import get_wsgi_application

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "settings")
application = get_wsgi_application()

import warnings

warnings.filterwarnings('ignore', message='Unverified HTTPS request')

from datetime import *

# print(f"inicio {datetime.now()}")

from helpdesk.models import *
from sagest import models
from django.db import transaction
from colorama import Back, init
from sga.models import *
from sga.funciones import variable_valor, notificacion2
from investigacion.models import *
import xlwt
import openpyxl
from xlwt import *
from settings import MEDIA_ROOT, BASE_DIR


def progress_bar(progress, total):
    percent = 100 * (progress / float(total))
    bar = (Back.GREEN + ' ' + Back.RESET) * int(percent) + '-' * (100 - int(percent))
    print(f"\r|{bar}| {percent:.2f}%", end="\r")


def corregir_secuencial(**kwargs):
    tables = []
    app = kwargs.pop('app',
                     'helpdesk')  # Si no se especifica el parametro 'app' al llamar a la funcion se usa 'helpdesk'
    cursor = connections['default'].cursor()
    sql = f"SELECT array_to_string(ARRAY(SELECT TABLE_NAME FROM information_schema.tables WHERE table_schema = 'public' AND table_name like '%{app}%'), ',') AS {app}_tables;"
    cursor.execute(sql)
    results = cursor.fetchall()
    if results: tables = results[0][0].split(',')
    for table in tables:
        try:
            sql = f"SELECT pg_get_serial_sequence('{table}', 'id');"
            cursor.execute(sql)
            serial_sequence = cursor.fetchall()[0][0]
            sql = f"SELECT SETVAL('{serial_sequence}', (select max(id) from {table}));"
            cursor.execute(sql)
            # print(f"Serial sequence of <<{table}>> updated.")
        except Exception as ex:
            print(ex.__str__())


@transaction.atomic()
def migrar_registros_helpdesk():
    try:
        val = None
        data = models.HdTipoIncidente.objects.all().values().order_by('id')
        instances = [HdTipoIncidente(**item) for item in data if
                     not HdTipoIncidente.objects.filter(pk=item['id']).exists()]
        HdTipoIncidente.objects.bulk_create(instances)

        for val in models.HdGrupo.objects.all().exclude(id__in=HdGrupo.objects.values_list('id', flat=True)).order_by(
                'id'):
            with transaction.atomic():
                try:
                    HdGrupo.objects.create(pk=val.pk, nombre=val.nombre, descripcion=val.descripcion,
                                           tipoincidente_id=val.tipoincidente.id, fecha_creacion=val.fecha_creacion,
                                           usuario_creacion=val.usuario_creacion,
                                           fecha_modificacion=val.fecha_modificacion,
                                           usuario_modificacion=val.usuario_modificacion)
                except Exception as ex:
                    linea_error = 'Error on line {}'.format(sys.exc_info()[-1].tb_lineno)
                    transaction.set_rollback(True)
                    print(f"ERROR: {ex.__str__()} - {linea_error}")

        for val in models.HdDetalle_Grupo.objects.all().exclude(
                id__in=HdDetalle_Grupo.objects.values_list('id', flat=True)).order_by('id'):
            with transaction.atomic():
                try:
                    HdDetalle_Grupo.objects.create(pk=val.pk, grupo_id=val.grupo.id, persona=val.persona,
                                                   responsable=val.responsable, estado=val.estado,
                                                   fecha_creacion=val.fecha_creacion,
                                                   usuario_creacion=val.usuario_creacion,
                                                   fecha_modificacion=val.fecha_modificacion,
                                                   usuario_modificacion=val.usuario_modificacion)
                except Exception as ex:
                    linea_error = 'Error on line {}'.format(sys.exc_info()[-1].tb_lineno)
                    transaction.set_rollback(True)
                    print(f"ERROR: {ex.__str__()} - {linea_error}")

        data = models.HdCategoria.objects.all().values().order_by('id')
        instances = [HdCategoria(**item) for item in data if not HdCategoria.objects.filter(pk=item['id']).exists()]
        HdCategoria.objects.bulk_create(instances)

        data = models.HdSubCategoria.objects.all().values().order_by('id')
        instances = [HdSubCategoria(**item) for item in data if
                     not HdSubCategoria.objects.filter(pk=item['id']).exists()]
        HdSubCategoria.objects.bulk_create(instances)

        data = models.HdUrgencia.objects.all().values().order_by('id')
        instances = [HdUrgencia(**item) for item in data if not HdUrgencia.objects.filter(pk=item['id']).exists()]
        HdUrgencia.objects.bulk_create(instances)

        data = models.HdImpacto.objects.all().values().order_by('id')
        instances = [HdImpacto(**item) for item in data if not HdImpacto.objects.filter(pk=item['id']).exists()]
        HdImpacto.objects.bulk_create(instances)

        data = models.HdPrioridad.objects.all().values().order_by('id')
        instances = [HdPrioridad(**item) for item in data if not HdPrioridad.objects.filter(pk=item['id']).exists()]
        HdPrioridad.objects.bulk_create(instances)

        data = models.HdUrgencia_Impacto_Prioridad.objects.all().values().order_by('id')
        instances = [HdUrgencia_Impacto_Prioridad(**item) for item in data if
                     not HdUrgencia_Impacto_Prioridad.objects.filter(pk=item['id']).exists()]
        HdUrgencia_Impacto_Prioridad.objects.bulk_create(instances)

        data = models.HdDetalle_SubCategoria.objects.all().values().order_by('id')
        instances = [HdDetalle_SubCategoria(**item) for item in data if
                     not HdDetalle_SubCategoria.objects.filter(pk=item['id']).exists()]
        HdDetalle_SubCategoria.objects.bulk_create(instances)

        data = models.HdMedioReporte.objects.all().values().order_by('id')
        instances = [HdMedioReporte(**item) for item in data if
                     not HdMedioReporte.objects.filter(pk=item['id']).exists()]
        HdMedioReporte.objects.bulk_create(instances)

        data = models.HdEstado.objects.all().values().order_by('id')
        instances = [HdEstado(**item) for item in data if not HdEstado.objects.filter(pk=item['id']).exists()]
        HdEstado.objects.bulk_create(instances)

        data = models.HdTipoIncidente.objects.all().values().order_by('id')
        instances = [HdTipoIncidente(**item) for item in data if
                     not HdTipoIncidente.objects.filter(pk=item['id']).exists()]
        HdTipoIncidente.objects.bulk_create(instances)

        data = models.HdCausas.objects.all().values().order_by('id')
        instances = [HdCausas(**item) for item in data if not HdCausas.objects.filter(pk=item['id']).exists()]
        HdCausas.objects.bulk_create(instances)

        data = models.OrdenTrabajo.objects.all().values().order_by('id')
        instances = [OrdenTrabajo(**item) for item in data if not OrdenTrabajo.objects.filter(pk=item['id']).exists()]
        OrdenTrabajo.objects.bulk_create(instances)

        cont, val = 0, None
        incidentes_sagest = models.HdIncidente.objects.filter(tipoincidente__id=2).exclude(
            pk__in=HdIncidente.objects.values_list('id', flat=True)).order_by('id')
        for val in incidentes_sagest:
            with transaction.atomic():
                try:
                    HdIncidente.objects.create(pk=val.pk,
                                               asunto=val.asunto,
                                               persona=val.persona,
                                               departamento=val.departamento,
                                               descripcion=val.descripcion,
                                               subcategoria_id=val.subcategoria.id if val.subcategoria else None,
                                               detallesubcategoria_id=val.detallesubcategoria.id if val.detallesubcategoria else None,
                                               activo=val.activo,
                                               fechareporte=val.fechareporte,
                                               horareporte=val.horareporte,
                                               medioreporte_id=val.medioreporte.id if val.medioreporte else None,
                                               director=val.director,
                                               archivo=val.archivo,
                                               estado_id=val.estado.id if val.estado else None,
                                               tipoincidente_id=val.tipoincidente.id if val.tipoincidente else None,
                                               ubicacion=val.ubicacion,
                                               causa_id=val.causa.id if val.causa else None,
                                               responsableactivofijo=val.responsableactivofijo,
                                               realizoencuesta=val.realizoencuesta,
                                               revisionequipoexterno=val.revisionequipoexterno,
                                               revisionequiposincodigo=val.revisionequiposincodigo,
                                               serie=val.serie,
                                               ordentrabajo_id=val.ordentrabajo.id if val.ordentrabajo else None,
                                               tercerapersona=val.tercerapersona,
                                               tipousuario=val.tipousuario,
                                               concodigo=val.concodigo,
                                               activosincodigo=val.activosincodigo,
                                               fecha_creacion=val.fecha_creacion,
                                               usuario_creacion=val.usuario_creacion,
                                               usuario_modificacion=val.usuario_modificacion,
                                               fecha_modificacion=val.fecha_modificacion,
                                               status=val.status)

                    cont += 1
                except Exception as ex:
                    linea_error = 'Error on line {}'.format(sys.exc_info()[-1].tb_lineno)
                    transaction.set_rollback(True)
                    print(f"ERROR: {ex.__str__()} - {linea_error}")

        proceso_sagest = models.HdProceso.objects.all().exclude(
            pk__in=HdProceso.objects.values_list('id', flat=True)).order_by('id')
        for val in proceso_sagest:
            with transaction.atomic():
                try:
                    if not HdProceso.objects.filter(pk=val.pk):
                        HdProceso.objects.create(id=val.pk, nombre=val.nombre, fecha_creacion=val.fecha_creacion,
                                                 usuario_creacion=val.usuario_creacion,
                                                 usuario_modificacion=val.usuario_modificacion,
                                                 fecha_modificacion=val.fecha_modificacion, status=val.status)
                except Exception as ex:
                    linea_error = 'Error on line {}'.format(sys.exc_info()[-1].tb_lineno)
                    transaction.set_rollback(True)
                    print(f"ERROR: {ex.__str__()} - {linea_error}")

        for val in models.HdEstado_Proceso.objects.all().exclude(
                id__in=HdEstado_Proceso.objects.values_list('id', flat=True)).order_by('id'):
            with transaction.atomic():
                try:
                    HdEstado_Proceso.objects.create(id=val.pk, nombre=val.nombre, proceso_id=val.proceso.id,
                                                    detalle=val.detalle, fecha_creacion=val.fecha_creacion,
                                                    usuario_creacion=val.usuario_creacion,
                                                    usuario_modificacion=val.usuario_modificacion,
                                                    fecha_modificacion=val.fecha_modificacion, status=val.status)
                except Exception as ex:
                    linea_error = 'Error on line {}'.format(sys.exc_info()[-1].tb_lineno)
                    transaction.set_rollback(True)
                    print(f"ERROR: {ex.__str__()} - {linea_error}")

        cont, val = 0, None
        detalle_sagest = models.HdDetalle_Incidente.objects.filter(
            Q(incidente__tipoincidente__id=2) | Q(incidente_id__isnull=True)).exclude(
            pk__in=HdDetalle_Incidente.objects.values_list('id', flat=True)).order_by('id')
        for val in detalle_sagest:
            with transaction.atomic():
                try:
                    HdDetalle_Incidente.objects.create(pk=val.pk,
                                                       incidente_id=val.incidente.id if val.incidente else None,
                                                       agente_id=val.agente.id if val.agente else None,
                                                       responsable=val.responsable,
                                                       grupo_id=val.grupo.id if val.grupo else None,
                                                       resolucion=val.resolucion,
                                                       fecharesolucion=val.fecharesolucion,
                                                       horaresolucion=val.horaresolucion,
                                                       estadoasignacion=val.estadoasignacion,
                                                       estadoproceso_id=val.estadoproceso.id if val.estadoproceso else None,
                                                       estado_id=val.estado.id if val.estado else None,
                                                       fecha_creacion=val.fecha_creacion,
                                                       usuario_creacion=val.usuario_creacion,
                                                       usuario_modificacion=val.usuario_modificacion,
                                                       fecha_modificacion=val.fecha_modificacion,
                                                       status=val.status)
                    cont += 1
                except Exception as ex:
                    linea_error = 'Error on line {}'.format(sys.exc_info()[-1].tb_lineno)
                    transaction.set_rollback(True)
                    print(f"ERROR: {ex.__str__()} - {linea_error}")

        data = models.HdDirector.objects.all().values().order_by('id')
        instances = [HdDirector(**item) for item in data if not HdDirector.objects.filter(pk=item['id']).exists()]
        HdDirector.objects.bulk_create(instances)

        corregir_secuencial()
    except Exception as ex:
        transaction.set_rollback(True)
        print('error: %s' % ex.__str__())


def migrar_tipoprofesor_sga_to_postulaciondip(**kwargs):
    try:
        from postulaciondip.models import TipoDocente, Convocatoria
        lista = list(Convocatoria.objects.values('id', 'tipodocente').all())

        with transaction.atomic():
            for tp in TipoProfesor.objects.values('id', 'nombre', 'abreviatura').all():
                if not TipoDocente.objects.values('id').filter(pk=tp['id']).exists():
                    try:
                        TipoDocente(**tp).save()
                    except Exception as ex:
                        linea_error = 'Error on line {}'.format(sys.exc_info()[-1].tb_lineno)
                        transaction.set_rollback(True)
                        print(f"ERROR: {ex.__str__()} - {linea_error}")

        with transaction.atomic():
            for cv in lista:
                convocatoria = Convocatoria.objects.get(id=cv['id'])
                if not convocatoria.tipodocente:
                    try:
                        convocatoria.tipodocente_id = cv['tipodocente']
                        convocatoria.save()
                    except Exception as ex:
                        linea_error = 'Error on line {}'.format(sys.exc_info()[-1].tb_lineno)
                        transaction.set_rollback(True)
                        print(f"ERROR: {ex.__str__()} - {linea_error}")

        corregir_secuencial(app='postulaciondip')
    except Exception as ex:
        linea_error = 'Error on line {}'.format(sys.exc_info()[-1].tb_lineno)
        print(f"ERROR: {ex.__str__()} - {linea_error}")


def barrido_por_estado_director_proyecto():
    try:
        parse_funcion_criterio = lambda x: {1: 55, 2: 56, 3: 57}[x] if x in (1, 2, 3) else None
        proyectos = ProyectoInvestigacion.objects.filter(status=True, estado=37).exclude(cerrado=True)
        count = 0
        for proyecto in proyectos:
            try:
                for evidencia in EvidenciaActividadDetalleDistributivo.objects.filter(
                        criterio__criterioinvestigacionperiodo__criterio__id=55,
                        criterio__distributivo__profesor=proyecto.profesor,
                        hasta__month__in=[5, 6, 7, 8, 9, 10, 11, 12], hasta__year=2023, estadoaprobacion__in=[2, 4],
                        status=True):
                    for integrante in proyecto.integrantes_proyecto().exclude(funcion=1):
                        if distributivo := DetalleDistributivo.objects.filter(
                                criterioinvestigacionperiodo__criterio__id=parse_funcion_criterio(integrante.funcion),
                                distributivo__profesor=integrante.profesor,
                                distributivo__periodo=evidencia.criterio.distributivo.periodo, status=True).first():
                            for _evidencia in EvidenciaActividadDetalleDistributivo.objects.filter(
                                    criterio=distributivo, hasta__month=evidencia.hasta.month,
                                    hasta__year=evidencia.hasta.year, status=True):
                                if not _evidencia.estadoaprobacion == evidencia.estadoaprobacion:
                                    _evidencia.desde = evidencia.desde
                                    _evidencia.hasta = evidencia.hasta
                                    _evidencia.actividad = evidencia.actividad
                                    _evidencia.aprobado = evidencia.aprobado
                                    _evidencia.archivo = evidencia.archivo
                                    _evidencia.usuarioaprobado = evidencia.usuarioaprobado
                                    _evidencia.fechaaprobado = evidencia.fechaaprobado
                                    _evidencia.estadoaprobacion = evidencia.estadoaprobacion
                                    _evidencia.archivofirmado = evidencia.archivofirmado
                                    _evidencia.save()

                                    _evidencia.anexoevidenciaactividad_set.filter(status=True).update(status=False)
                                    for anexo in evidencia.anexoevidenciaactividad_set.filter(
                                        status=True): AnexoEvidenciaActividad(evidencia=_evidencia,
                                                                              observacion=anexo.observacion,
                                                                              archivo=anexo.archivo).save()

                                    _evidencia.evidenciaactividadaudi_set.filter(status=True).update(status=False)
                                    for anexo in evidencia.evidenciaactividadaudi_set.filter(
                                        status=True): EvidenciaActividadAudi(evidencia=_evidencia,
                                                                             archivo=anexo.archivo).save()

                                    if h := HistorialAprobacionEvidenciaActividad.objects.filter(evidencia=evidencia,
                                                                                                 estadoaprobacion__in=[
                                                                                                     2, 4],
                                                                                                 status=True).order_by(
                                            '-id').first():
                                        HistorialAprobacionEvidenciaActividad(evidencia=_evidencia,
                                                                              aprobacionpersona=h.aprobacionpersona,
                                                                              observacion=h.observacion,
                                                                              fechaaprobacion=h.fechaaprobacion,
                                                                              estadoaprobacion=h.estadoaprobacion).save()

                                    count += 1
                                    print(
                                        f"{count}.- INFORME DEL MES {_evidencia.hasta.month} - {integrante.persona} migrado correctamente... {evidencia.estadoaprobacion}/{_evidencia.estadoaprobacion}")
            except Exception as ex:
                print(f"Error en la migración del proyecto {proyecto.pk} {proyecto}. Exception {ex.__str__()}")
    except Exception as ex:
        pass


def eliminar_evidencia():
    try:
        print('Eliminando registros...')
        for evidencia in EvidenciaActividadDetalleDistributivo.objects.filter(
                criterio__criteriodocenciaperiodo__criterio__id=167, criterio__criteriodocenciaperiodo__periodo__id=177,
                hasta__month__in=[7, 8], desde__month__in=[7, 8], desde__year=2023, hasta__year=2023, status=True):
            print(
                f"{evidencia.criterio.distributivo} - {evidencia.criterio.criteriodocenciaperiodo.criterio} eliminado...")
            evidencia.delete()
    except Exception as ex:
        pass


def barrido_por_evidencia_director_proyecto_investigacion():
    try:
        parse_funcion_criterio = lambda x: {1: 55, 2: 56, 3: 57}[x] if x in (1, 2, 3) else None
        _count0 = 0
        for evidencia in EvidenciaActividadDetalleDistributivo.objects.filter(
                pk__in=variable_valor('ID_EVIDENCIA_PROYECTO_INVESTIGACION')):
            _count0 += 1
            try:
                if proyecto := ProyectoInvestigacion.objects.filter(profesor=evidencia.criterio.distributivo.profesor,
                                                                    estado__id=37, status=True).exclude(
                        cerrado=True).first():
                    _count = 0
                    for integrante in proyecto.integrantes_proyecto().exclude(funcion=1):
                        _count += 1
                        try:
                            _evidencia = None
                            if distributivo := DetalleDistributivo.objects.filter(
                                    criterioinvestigacionperiodo__criterio__id=parse_funcion_criterio(
                                            integrante.funcion), distributivo__profesor=integrante.profesor,
                                    distributivo__periodo=evidencia.criterio.distributivo.periodo, status=True).first():
                                _evidencia = EvidenciaActividadDetalleDistributivo.objects.filter(criterio=distributivo,
                                                                                                  hasta__month=evidencia.hasta.month,
                                                                                                  hasta__year=evidencia.hasta.year,
                                                                                                  status=True).first()
                                if _evidencia:
                                    _evidencia.desde = evidencia.desde
                                    _evidencia.hasta = evidencia.hasta
                                    _evidencia.actividad = evidencia.actividad
                                    _evidencia.aprobado = evidencia.aprobado
                                    _evidencia.archivo = evidencia.archivo
                                    _evidencia.usuarioaprobado = evidencia.usuarioaprobado
                                    _evidencia.fechaaprobado = evidencia.fechaaprobado
                                    _evidencia.estadoaprobacion = evidencia.estadoaprobacion
                                    _evidencia.archivofirmado = evidencia.archivofirmado
                                else:
                                    _evidencia = EvidenciaActividadDetalleDistributivo(desde=evidencia.desde,
                                                                                       hasta=evidencia.hasta,
                                                                                       actividad=evidencia.actividad,
                                                                                       aprobado=evidencia.aprobado,
                                                                                       archivo=evidencia.archivo,
                                                                                       usuarioaprobado=evidencia.usuarioaprobado,
                                                                                       fechaaprobado=evidencia.fechaaprobado,
                                                                                       estadoaprobacion=evidencia.estadoaprobacion,
                                                                                       archivofirmado=evidencia.archivofirmado,
                                                                                       criterio=distributivo)
                                _evidencia.save()
                                _evidencia.anexoevidenciaactividad_set.filter(status=True).delete()
                                for anexo in evidencia.anexoevidenciaactividad_set.filter(
                                    status=True): AnexoEvidenciaActividad(evidencia=_evidencia,
                                                                          observacion=anexo.observacion,
                                                                          archivo=anexo.archivo,
                                                                          fecha_creacion=anexo.fecha_creacion).save()
                                _evidencia.evidenciaactividadaudi_set.filter(status=True).delete()
                                for anexo in evidencia.evidenciaactividadaudi_set.filter(
                                    status=True): EvidenciaActividadAudi(evidencia=_evidencia, archivo=anexo.archivo,
                                                                         fecha_creacion=anexo.fecha_creacion).save()
                                _evidencia.historialaprobacionevidenciaactividad_set.filter(status=True).delete()
                                for anexo in evidencia.historialaprobacionevidenciaactividad_set.filter(
                                    status=True): HistorialAprobacionEvidenciaActividad(evidencia=_evidencia,
                                                                                        fecha_creacion=anexo.fecha_creacion,
                                                                                        aprobacionpersona=anexo.aprobacionpersona,
                                                                                        observacion=anexo.observacion,
                                                                                        fechaaprobacion=anexo.fechaaprobacion,
                                                                                        estadoaprobacion=anexo.estadoaprobacion).save()
                                print(
                                    f"Evidencia {_count0}.- {integrante.persona} migrado correctamente. {_count}/{proyecto.integrantes_proyecto().exclude(funcion=1).values('id').count() - 1}")
                        except Exception as ex1:
                            print(f"{ex1.__str__()}")
            except Exception as ex2:
                print(f"Error en la migración de la evidencia {evidencia.pk} {evidencia}. Exception {ex2.__str__()}")
    except Exception as ex3:
        print(f"{ex3.__str__()}")


def actualiza_estado_evidencia_criterios_varios():
    try:
        from sga.models import CriterioDocencia, CriterioInvestigacion, CriterioVinculacion

        CRITERIOS_DOCENCIA = [131, 166, 132, 19, 133, 168]
        CRITERIOS_INVESTIGACION = [59]
        CRITERIOS_VINCULACION = [4]

        # for x in CriterioDocencia.objects.filter(pk__in=CRITERIOS_DOCENCIA): print(x.nombre)
        # print(CriterioInvestigacion.objects.filter(pk__in=CRITERIOS_INVESTIGACION))

        for evidencia in EvidenciaActividadDetalleDistributivo.objects.filter(
                Q(Q(criterio__criteriodocenciaperiodo__criterio__id__in=CRITERIOS_DOCENCIA) |
                  Q(criterio__criterioinvestigacionperiodo__criterio__id__in=CRITERIOS_INVESTIGACION) |
                  Q(criterio__criteriovinculacionperiodo__criterio__id__in=CRITERIOS_VINCULACION) |
                  Q(criterio__criteriogestionperiodo__isnull=False)) &
                Q(estadoaprobacion__in=(1, 4),
                  desde__year=2023, hasta__year=2023,
                  criterio__distributivo__periodo__id=177,
                  hasta__month__in=[4, 5, 6, 7, 8], status=True)):
            if not evidencia.archivofirmado:
                if criterio := evidencia.criterio.criteriodocenciaperiodo:
                    print(
                        f"{evidencia.pk}.- MES {evidencia.hasta.month}, {evidencia.criterio.distributivo.profesor} - {criterio.criterio.nombre}")
                if criterio := evidencia.criterio.criterioinvestigacionperiodo:
                    print(
                        f"{evidencia.pk}.- MES {evidencia.hasta.month}, {evidencia.criterio.distributivo.profesor} - {criterio.criterio.nombre}")
                if criterio := evidencia.criterio.criteriovinculacionperiodo:
                    print(
                        f"{evidencia.pk}.- MES {evidencia.hasta.month}, {evidencia.criterio.distributivo.profesor} - {criterio.criterio.nombre}")
                if criterio := evidencia.criterio.criteriogestionperiodo:
                    print(
                        f"{evidencia.pk}.- MES {evidencia.hasta.month}, {evidencia.criterio.distributivo.profesor} - {criterio.criterio.nombre}")
                evidencia.estadoaprobacion = 2
                evidencia.save()
    except Exception as ex:
        pass


def reporte_porcentaje_cumplimiento():
    periodo = Periodo.objects.get(pk=177)
    # profesor = Profesor.objects.get(persona__cedula='0919615906') # pk=278

    print(periodo.__str__())
    # print(profesor.__str__())

    try:
        response = []
        distributivos = ProfesorDistributivoHoras.objects.filter(periodo=periodo, activo=True, status=True)
        _count, forloop = distributivos.count(), 0
        for distributivo in distributivos:
            profesor = distributivo.profesor
            d = distributivo.profesor.informe_actividades_mensual_docente_v4(periodo, '01-08-2023', '31-08-2023',
                                                                             'FACULTAD')
            fini, ffin, asignaturas = date(2023, 8, 1), date(2023, 8, 31), d.get('asignaturas')
            criteriodocenciaperiodo = d.get('criteriodocenciaperiodo')
            _data = []
            subtotal, total, numeroactividades = 0, 0, 0
            forloop += 1
            for actividad in distributivo.detalle_horas_docencia(fini, ffin):
                flag, _data = False, {}
                if actividad.criteriodocenciaperiodo:
                    if actividad.criteriodocenciaperiodo.htmldocente:
                        if actividad.criteriodocenciaperiodo.nombrehtmldocente() == 'actividaddocente':
                            if actividaddocente := actividad.criteriodocenciaperiodo.horarios_actividaddocente_profesor(
                                    profesor, fini, ffin):
                                total += 100
                                flag = True

                    if actividad.criteriodocenciaperiodo.nombrehtmldocente() == 'impartirclase':
                        totalimpartir = actividad.criteriodocenciaperiodo.totalimparticlase(distributivo.profesor, fini,
                                                                                            ffin, asignaturas)
                        if totalimpartir[0][0]:
                            _value = totalimpartir[0][2]
                            total += _value
                            if _value < 100: _data = {'tipo': 1, 'criterio': actividad.criteriodocenciaperiodo,
                                                      'porcentaje': _value}
                            flag = True

                    # Evidencia Moodle
                    if actividad.criteriodocenciaperiodo.nombrehtmldocente() == 'evidenciamoodle':
                        if listadoevidencias := actividad.criteriodocenciaperiodo.horario_evidencia_moodle(profesor,
                                                                                                           d.get('finicresta'),
                                                                                                           d.get('ffincresta')):
                            evidencia = listadoevidencias[-1]
                            if not evidencia[11] == 4:
                                _value = evidencia[10]
                                total += _value
                                if evidencia[10] < 100: _data = {'tipo': 1,
                                                                 'criterio': actividad.criteriodocenciaperiodo,
                                                                 'porcentaje': _value}
                                flag = True

                    # Actividades Moodle
                    if actividad.criteriodocenciaperiodo.nombrehtmldocente() == 'materialsilabo':
                        if actividadhor := actividad.criteriodocenciaperiodo.horarios_actividad_profesor(profesor, fini,
                                                                                                         ffin):
                            if actividadhor[-1][3]:
                                _value = actividadhor[-1][3]
                                total += _value
                                if _value < 100: _data = {'tipo': 1, 'criterio': actividad.criteriodocenciaperiodo,
                                                          'porcentaje': _value}
                                flag = True

                    if actividad.criteriodocenciaperiodo.nombrehtmldocente() == 'cursonivelacion':
                        if actividadnivelacioncarrera := actividad.criteriodocenciaperiodo.horarios_nivelacioncarrera_profesor(
                                profesor, fini, ffin):
                            total += 100
                            flag = True

                    if actividad.criteriodocenciaperiodo.nombrehtmldocente() == 'planificarcontenido':
                        if contenidohor := actividad.criteriodocenciaperiodo.horarios_contenido_profesor(profesor, fini,
                                                                                                         ffin):
                            evidencia = contenidohor[-1]
                            _value = evidencia[3]
                            total += _value
                            if _value < 100: _data = {'tipo': 1, 'criterio': actividad.criteriodocenciaperiodo,
                                                      'porcentaje': _value}
                            flag = True

                    if actividad.criteriodocenciaperiodo.nombrehtmldocente() == 'tutoriaacademica':
                        if tutoriasacademicas := actividad.criteriodocenciaperiodo.horarios_tutoriasacademicas_profesor(
                                profesor, fini, ffin):
                            _value = tutoriasacademicas[0][3]
                            total += _value
                            if _value < 100: _data = {'tipo': 1, 'criterio': actividad.criteriodocenciaperiodo,
                                                      'porcentaje': _value}
                            flag = True

                    if actividad.criteriodocenciaperiodo.nombrehtmldocente() == 'seguimientoplataforma':
                        if listadoseguimientos := actividad.criteriodocenciaperiodo.horario_seguimiento_tutor_fecha(
                                profesor, fini, ffin):
                            s = listadoseguimientos[-1]
                            _value = s[9]
                            total += _value
                            if _value < 100: _data = {'tipo': 1, 'criterio': actividad.criteriodocenciaperiodo,
                                                      'porcentaje': _value}
                            flag = True

                    if actividad.criteriodocenciaperiodo.nombrehtmldocente() == 'nivelacioncarrera':
                        if actividadgestion := actividad.criteriodocenciaperiodo.horarios_informesdocencia_profesor(
                                distributivo, fini, ffin):
                            if actigestion := actividadgestion.listadoevidencias:
                                if temp := list(filter(lambda x: x[0] == 2, actigestion)):
                                    _value = temp[0][2]
                                    total += _value
                                    if _value < 100: _data = {'tipo': 1, 'criterio': actividad.criteriodocenciaperiodo,
                                                              'porcentaje': _value}
                                    flag = True

                    if actividad.criteriodocenciaperiodo.nombrehtmldocente() == 'seguimientotransversal':
                        if listadoseguimientos := actividad.criteriodocenciaperiodo.horario_seguimiento_transaversal_fecha(
                                profesor, fini, ffin):
                            s = listadoseguimientos[-1]
                            _value = s[9]
                            total += _value
                            if _value < 100: _data = {'tipo': 1, 'criterio': actividad.criteriodocenciaperiodo,
                                                      'porcentaje': _value}
                            flag = True

                    if actividad.criteriodocenciaperiodo.nombrehtmldocente() == 'apoyovicerrectorado':
                        if actividadapoyo := actividad.criteriodocenciaperiodo.horarios_apoyo_profesor(profesor, fini,
                                                                                                       ffin):
                            total += 100
                            flag = True

                if flag:
                    numeroactividades += 1
                    # print(f"DOC - {numeroactividades}.- {'%.2f' % total}  {actividad.criteriodocenciaperiodo.criterio}")
                    subtotal = total

            for actividad in distributivo.detalle_horas_investigacion():
                flag = False
                if actividad.criterioinvestigacionperiodo:
                    if actividad.criterioinvestigacionperiodo.nombrehtmldocente() == 'actividadinvestigacion':
                        if actividadgestion := actividad.criterioinvestigacionperiodo.horarios_informesinvestigacion_profesor(
                                distributivo, fini, ffin):
                            if actividadgestion.listadoevidencias:
                                for actigestion in actividadgestion.listadoevidencias:
                                    if actigestion[0] == 2:
                                        _value = actigestion[2]
                                        total += _value
                                        if _value < 100: _data = {'tipo': 2,
                                                                  'criterio': actividad.criterioinvestigacionperiodo,
                                                                  'porcentaje': _value}
                                        flag = True

                if flag:
                    numeroactividades += 1
                    # print(f"INV - {numeroactividades}.- {'%.2f' % total}  {actividad.criterioinvestigacionperiodo.criterio}")
                    subtotal = total

            for actividad in distributivo.detalle_horas_gestion(fini, ffin):
                flag = False
                if actividad.criteriogestionperiodo:
                    if actividad.criteriogestionperiodo.nombrehtmldocente() == 'actividadgestion':
                        if actividadgestion := actividad.criteriogestionperiodo.horarios_actividadgestion_profesor(
                                profesor, fini, ffin):
                            total += 100
                            flag = True

                if flag:
                    numeroactividades += 1
                    # print(f"GES - {numeroactividades}.- {'%.2f' % total}  {actividad.criteriogestionperiodo.criterio}")
                    subtotal = total

            for actividad in distributivo.detalle_horas_vinculacion():
                flag = False
                if actividad.criteriodocenciaperiodo.nombrehtmldocente() == 'actividadvinculacion':
                    if actividadgestion := actividad.criteriodocenciaperiodo.horarios_informesdocencia_profesor(
                            distributivo, fini, ffin):
                        if actividadgestion.listadoevidencias:
                            for actigestion in actividadgestion.listadoevidencias:
                                if actigestion[0] == 2:
                                    _value = actigestion[2]
                                    total += _value
                                    if _value < 100: _data = {'tipo': 4, 'criterio': actividad.criteriodocenciaperiodo,
                                                              'porcentaje': _value}
                                    flag = True

                if flag:
                    numeroactividades += 1
                    # print(f"VIN - {numeroactividades}.- {total}  {actividad.criteriodocenciaperiodo.criterio}")
                    subtotal = total

            _porcent = 0

            try:
                _porcent = total / numeroactividades
            except ZeroDivisionError as ex:
                pass

            response.append([profesor, _porcent, _data])
            print([profesor, _porcent, _data])
        return response
    except Exception as ex:
        print(ex.__str__())


def get_promedio():
    try:
        import calendar
        periodo = Periodo.objects.get(pk=177)
        distributivo = ProfesorDistributivoHoras.objects.filter(profesor__persona__usuario__username='psalasp',
                                                                periodo=periodo, activo=True, status=True).first()
        now = datetime.now().date()
        fini, ffin = date(now.year, 8, 1), date(now.year, 8, calendar.monthrange(now.year, 8)[1])
        profesor = distributivo.profesor
        _data = []
        subtotal, total, numeroactividades = 0, 0, 0
        finicresta = fini - timedelta(days=5)
        ffincresta = ffin - timedelta(days=5)

        if distributivo.detalledistributivo_set.values('id').filter(criteriodocenciaperiodo__criterio__tipo=1,
                                                                    status=True).exists():
            for actividad in distributivo.detalle_horas_docencia(fini, ffin):
                if h := actividad.actividaddetalledistributivofecha(fini, ffin):
                    if not h.horas:
                        continue
                flag, html = False, actividad.criteriodocenciaperiodo.nombrehtmldocente()
                if actividad.criteriodocenciaperiodo:
                    if actividad.criteriodocenciaperiodo.htmldocente():
                        if html == 'actividaddocente':
                            if actividaddocente := actividad.criteriodocenciaperiodo.horarios_actividaddocente_profesor(
                                    profesor, fini, ffin):
                                if actividaddocente.totalplanificadas:
                                    total += 100
                                    flag = True

                    if html == 'impartirclase':
                        profesormateria = ProfesorMateria.objects.filter(profesor=profesor,
                                                                         materia__nivel__periodo=periodo,
                                                                         tipoprofesor__imparteclase=True, activo=True,
                                                                         materia__status=True).distinct().order_by(
                            'desde', 'materia__asignatura__nombre')
                        if periodo.clasificacion == 1:
                            asignaturas = profesormateria.filter(((Q(desde__gte=fini) & Q(hasta__lte=ffin)) | (
                                        Q(desde__lte=fini) & Q(hasta__gte=ffin)) | (Q(desde__lte=ffin) & Q(
                                desde__gte=fini)) | (Q(hasta__gte=fini) & Q(hasta__lte=ffin)))).distinct().exclude(
                                tipoprofesor_id=15, materia__modeloevaluativo_id__in=[27]).order_by('desde',
                                                                                                    'materia__asignatura__nombre')
                        else:
                            asignaturas = profesormateria.filter(((Q(desde__gte=fini) & Q(hasta__lte=ffin)) | (
                                        Q(desde__lte=fini) & Q(hasta__gte=ffin)) | (Q(desde__lte=ffin) & Q(
                                desde__gte=fini)) | (Q(hasta__gte=fini) & Q(hasta__lte=ffin)))).distinct().order_by(
                                'desde', 'materia__asignatura__nombre')

                        totalimpartir = actividad.criteriodocenciaperiodo.totalimparticlase(distributivo.profesor, fini,
                                                                                            ffin, asignaturas)

                        if totalimpartir[0][0]:
                            _value = totalimpartir[0][2]
                            total += _value
                            if _value < 100: _data.append(
                                {'tipo': 1, 'criterio': actividad.criteriodocenciaperiodo, 'porcentaje': _value})
                            flag = True

                    if html == 'evidenciamoodle':
                        if listadoevidencias := actividad.criteriodocenciaperiodo.horario_evidencia_moodle(profesor,
                                                                                                           finicresta,
                                                                                                           ffincresta):
                            evidencia = listadoevidencias[-1]
                            if not evidencia[11] == 4 and evidencia[1]:
                                _value = evidencia[10]
                                total += _value
                                if evidencia[10] < 100: _data.append(
                                    {'tipo': 1, 'criterio': actividad.criteriodocenciaperiodo, 'porcentaje': _value})
                                flag = True

                    if html == 'materialsilabo':
                        if actividadhor := actividad.criteriodocenciaperiodo.horarios_actividad_profesor(profesor, fini,
                                                                                                         ffin):
                            if not actividadhor.__len__() <= 2:
                                _value = 0
                                for acti in actividadhor:
                                    if not acti[4] == 1:
                                        _value = acti[3]
                                if _value:
                                    total += _value
                                    if _value < 100: _data.append(
                                        {'tipo': 1, 'criterio': actividad.criteriodocenciaperiodo,
                                         'porcentaje': _value})
                                    flag = True

                    if html == 'cursonivelacion':
                        if actividadnivelacioncarrera := actividad.criteriodocenciaperiodo.horarios_nivelacioncarrera_profesor(
                                profesor, fini, ffin):
                            total += 100

                            flag = True

                    if html == 'planificarcontenido':
                        if contenidohor := actividad.criteriodocenciaperiodo.horarios_contenido_profesor(profesor, fini,
                                                                                                         ffin):
                            _value = 0
                            for x in contenidohor:
                                if not x[6] == 3 and x[4]: _value = x[3]
                            if _value:
                                total += _value
                                if _value < 100: _data.append(
                                    {'tipo': 1, 'criterio': actividad.criteriodocenciaperiodo, 'porcentaje': _value})
                                flag = True

                    if html == 'tutoriaacademica':
                        if tutoriasacademicas := actividad.criteriodocenciaperiodo.horarios_tutoriasacademicas_profesor(
                                profesor, fini, ffin):
                            if tutoriasacademicas[0][1]:
                                _value = tutoriasacademicas[0][3]
                                total += _value
                                if _value < 100: _data.append(
                                    {'tipo': 1, 'criterio': actividad.criteriodocenciaperiodo, 'porcentaje': _value})
                                flag = True

                    if html == 'seguimientoplataforma':
                        if listadoseguimientos := actividad.criteriodocenciaperiodo.horario_seguimiento_tutor_fecha(
                                profesor, fini, ffin):
                            s = listadoseguimientos[-1]
                            if s[11]:
                                _value = s[9]
                                total += _value
                                if _value < 100: _data.append(
                                    {'tipo': 1, 'criterio': actividad.criteriodocenciaperiodo, 'porcentaje': _value})
                                flag = True

                    if html == 'nivelacioncarrera':
                        if actividadgestion := actividad.criteriodocenciaperiodo.horarios_informesdocencia_profesor(
                                distributivo, fini, ffin):
                            if actividadgestion.totalplanificadas:
                                if actigestion := actividadgestion.listadoevidencias:
                                    if temp := list(filter(lambda x: x[0] == 2, actigestion)):
                                        _value = temp[0][2]
                                        total += _value
                                        if _value < 100: _data.append(
                                            {'tipo': 1, 'criterio': actividad.criteriodocenciaperiodo,
                                             'porcentaje': _value})
                                        flag = True

                    if html == 'seguimientotransversal':
                        if listadoseguimientos := actividad.criteriodocenciaperiodo.horario_seguimiento_transaversal_fecha(
                                profesor, fini, ffin):
                            s = listadoseguimientos[-1]
                            if s[11]:
                                _value = s[9]
                                total += _value
                                if _value < 100: _data.append(
                                    {'tipo': 1, 'criterio': actividad.criteriodocenciaperiodo, 'porcentaje': _value})
                                flag = True

                    if html == 'apoyovicerrectorado':
                        if actividadapoyo := actividad.criteriodocenciaperiodo.horarios_apoyo_profesor(profesor, fini,
                                                                                                       ffin):
                            if actividadapoyo.totalplanificadas:
                                total += 100
                                flag = True

                if flag:
                    numeroactividades += 1
                    subtotal = total

        if distributivo.detalledistributivo_set.values('id').filter(
                criterioinvestigacionperiodo__isnull=False).exists():
            for actividad in distributivo.detalle_horas_investigacion():
                flag = False
                if actividad.criterioinvestigacionperiodo:
                    if actividad.criterioinvestigacionperiodo.nombrehtmldocente() == 'actividadinvestigacion':
                        if actividadgestion := actividad.criterioinvestigacionperiodo.horarios_informesinvestigacion_profesor(
                                distributivo, fini, ffin):
                            if actividadgestion.listadoevidencias and actividadgestion[0][1]:
                                for actigestion in actividadgestion.listadoevidencias:
                                    if actividadgestion.totalplanificadas:
                                        if actigestion[0] == 2:
                                            _value = actigestion[2]
                                            total += _value
                                            if _value < 100: _data.append(
                                                {'tipo': 2, 'criterio': actividad.criterioinvestigacionperiodo,
                                                 'porcentaje': _value})
                                            flag = True

                if flag:
                    numeroactividades += 1
                    subtotal = total

        if distributivo.detalledistributivo_set.values('id').filter(criteriogestionperiodo__isnull=False).exists():
            for actividad in distributivo.detalle_horas_gestion(fini, ffin):
                flag = False
                if actividad.criteriogestionperiodo:
                    if actividad.criteriogestionperiodo.nombrehtmldocente() == 'actividadgestion':
                        if actividadgestion := actividad.criteriogestionperiodo.horarios_actividadgestion_profesor(
                                profesor, fini, ffin):
                            if actividadgestion.totalplanificadas:
                                total += 100
                                flag = True

                if flag:
                    numeroactividades += 1
                    subtotal = total

        if distributivo.detalledistributivo_set.values('id').filter(criteriodocenciaperiodo__isnull=False,
                                                                    criteriodocenciaperiodo__criterio__tipo=2).exists():
            for actividad in distributivo.detalle_horas_vinculacion():
                flag = False
                if actividad.criteriodocenciaperiodo.nombrehtmldocente() == 'actividadvinculacion':
                    if actividadgestion := actividad.criteriodocenciaperiodo.horarios_informesdocencia_profesor(
                            distributivo, fini, ffin):
                        if actividadgestion.listadoevidencias and actividadgestion.totalplanificadas:
                            for actigestion in actividadgestion.listadoevidencias:
                                if actigestion[0] == 2:
                                    _value = actigestion[2]
                                    total += _value
                                    if _value < 100: _data.append(
                                        {'tipo': 4, 'criterio': actividad.criteriodocenciaperiodo,
                                         'porcentaje': _value})
                                    flag = True

                if flag:
                    numeroactividades += 1
                    subtotal = total

        _porcent = 0
        try:
            _porcent = total / numeroactividades
        except ZeroDivisionError as ex:
            pass

        return [profesor, _porcent, _data, distributivo.carrera]
    except Exception as ex:
        pass


def reporte_docentes_admision():
    coordinacion = Coordinacion.objects.get(pk=9)
    periodo = Periodo.objects.get(pk=224)
    _ids_materia = []
    for nivel in Nivel.objects.filter(nivellibrecoordinacion__coordinacion=coordinacion, periodo=periodo):
        _ids_materia += nivel.materia_set.filter(status=True).values_list('id', flat=True).distinct()

    exclude = Coordinacion.objects.filter(status=True).values_list('id', flat=True).exclude(pk=9)
    _exclude = Materia.objects.filter(nivel__nivellibrecoordinacion__coordinacion__in=exclude,
                                      nivel__periodo=periodo).values_list('id', flat=True)

    _ids = ProfesorMateria.objects.values_list('profesor', flat=True).filter(materia__in=_ids_materia,
                                                                             materia__nivel__periodo=periodo).exclude(
        materia__id__in=_exclude)
    profesores = Profesor.objects.filter(id__in=_ids, status=True)
    for i, pm in enumerate(profesores):
        print("%s %s" % (i + 1, pm))


def migrar_evidencia_director_grupo_investigacion():
    from investigacion.models import GrupoInvestigacion, GrupoInvestigacionIntegrante
    from sga.models import EvidenciaActividadDetalleDistributivo
    from sga.templatetags.sga_extras import nombremes
    from django.http import HttpResponse

    CRITERIO_INTEGRANTE_DIRECTOR_GRUPO_INVESTIGACION = 58
    MONTH = 9

    try:

        criterio = CriterioInvestigacion.objects.get(pk=CRITERIO_INTEGRANTE_DIRECTOR_GRUPO_INVESTIGACION)
        os.makedirs(os.path.join(BASE_DIR, 'media'), exist_ok=True)
        os.makedirs(os.path.join(BASE_DIR, 'media', 'backup_arreglo_jefferson'), exist_ok=True)

        now = datetime.now()

        nombre_archivo = f'reporte_evidencias_eliminadas_{now.strftime("%Y%m%d_%H%M%S")}.xls'
        directory = os.path.join(os.path.join(BASE_DIR, 'media', 'backup_arreglo_jefferson'), nombre_archivo)

        __author__ = 'Unemi'
        fuentenormal = easyxf(
            'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
        titulo2 = easyxf(
            'font: name Verdana, color-index black, bold on , height 250; alignment: horiz centre;borders: left thin, right thin, top thin, bottom thin')
        encabesado_tabla = easyxf(
            'font: name Verdana , bold on , height 150; alignment: horiz left;borders: left thin, right thin, top thin, bottom thin;pattern: pattern solid, fore_colour gray25')
        encabesado_tabla_center = easyxf(
            'font: name Verdana , bold on , height 150; alignment: horiz centre;borders: left thin, right thin, top thin, bottom thin;pattern: pattern solid, fore_colour gray25')

        font_style = XFStyle()
        font_style.font.bold = True
        font_style2 = XFStyle()
        font_style2.font.bold = False
        wb = Workbook(encoding='utf-8')
        ws = wb.add_sheet('PORCENTAJES')
        ws.write_merge(0, 0, 0, 4, 'LISTADO DE EVIDENCIAS ELIMINADAS EN EL MES DE %s' % nombremes(MONTH).upper(),
                       titulo2)
        ws.write_merge(1, 1, 0, 4, '%s' % criterio.nombre, titulo2)
        response = HttpResponse(content_type="application/ms-excel")
        response['Content-Disposition'] = f'attachment; filename={nombre_archivo}'

        ws.col(0).width = 2000
        ws.col(1).width = 20500
        ws.col(2).width = 20500
        ws.col(3).width = 10000
        ws.col(4).width = 20500

        row_num = 2

        ws.write(row_num, 0, "Nº", encabesado_tabla)
        ws.write(row_num, 1, "DOCENTE", encabesado_tabla)
        ws.write(row_num, 2, "ACTIVIDAD", encabesado_tabla)
        ws.write(row_num, 3, "FECHA ELIMINACION", encabesado_tabla)
        ws.write(row_num, 4, "PERIODO", encabesado_tabla)

        date_format = xlwt.XFStyle()
        date_format.num_format_str = 'yyyy/mm/dd'
        c = 1
        row_num += 1

        for grupo in GrupoInvestigacion.objects.filter(vigente=True, status=True):
            if director := grupo.director():
                if detalledistributivo := DetalleDistributivo.objects.filter(
                        criterioinvestigacionperiodo__criterio__id=CRITERIO_INTEGRANTE_DIRECTOR_GRUPO_INVESTIGACION,
                        distributivo__profesor__persona=director.persona, distributivo__periodo_id=224,
                        distributivo__activo=True, status=True).first():
                    if evidencia := EvidenciaActividadDetalleDistributivo.objects.filter(criterio=detalledistributivo,
                                                                                         hasta__month=MONTH,
                                                                                         hasta__year=2023,
                                                                                         status=True).first():
                        evidencia.grupoinvestigacion = grupo
                        evidencia.save()

                        # Reporte eliminados
                        try:
                            for integrante in grupo.grupoinvestigacionintegrante_set.filter(status=True).exclude(
                                    funcion=1):
                                distributivo = DetalleDistributivo.objects.filter(
                                    criterioinvestigacionperiodo__criterio__id=CRITERIO_INTEGRANTE_DIRECTOR_GRUPO_INVESTIGACION,
                                    distributivo__profesor__persona=integrante.persona,
                                    distributivo__periodo=detalledistributivo.distributivo.periodo,
                                    distributivo__activo=True, status=True).first()
                                if distributivo:

                                    # Eliminacion de evidencias pre cargadas
                                    for e in EvidenciaActividadDetalleDistributivo.objects.filter(criterio=distributivo,
                                                                                                  hasta__month=evidencia.hasta.month,
                                                                                                  hasta__year=evidencia.hasta.year,
                                                                                                  status=True):
                                        ws.write(row_num, 0, f"{c}", fuentenormal)
                                        ws.write(row_num, 1, f"{e.criterio.distributivo.profesor.persona}",
                                                 fuentenormal)
                                        ws.write(row_num, 2, f"{e.actividad}", fuentenormal)
                                        ws.write(row_num, 3, f"{now.strftime('%Y%m%d_%H%M%S')}", fuentenormal)
                                        ws.write(row_num, 4, f"{evidencia.criterio.distributivo.periodo.nombre}",
                                                 fuentenormal)
                                        row_num += 1
                                        c += 1

                                        e.delete()

                                    wb.save(directory)
                                    # --------------------------------------------------

                                    _evidencia = grupo.evidenciaactividaddetalledistributivo_set.filter(
                                        criterio=distributivo, hasta__month=evidencia.hasta.month,
                                        hasta__year=evidencia.hasta.year, status=True).first()
                                    if _evidencia:
                                        _evidencia.actividaddetalledistributivo = evidencia.actividaddetalledistributivo
                                        _evidencia.desde = evidencia.desde
                                        _evidencia.hasta = evidencia.hasta
                                        _evidencia.actividad = evidencia.actividad
                                        _evidencia.aprobado = evidencia.aprobado
                                        _evidencia.archivo = evidencia.archivo
                                        _evidencia.usuarioaprobado = evidencia.usuarioaprobado
                                        _evidencia.fechaaprobado = evidencia.fechaaprobado
                                        _evidencia.estadoaprobacion = evidencia.estadoaprobacion
                                        _evidencia.archivofirmado = evidencia.archivofirmado
                                        _evidencia.grupoinvestigacion = evidencia.grupoinvestigacion
                                    else:
                                        _dict = \
                                        EvidenciaActividadDetalleDistributivo.objects.filter(pk=evidencia.pk).values()[
                                            0]
                                        _dict.pop('id')
                                        _evidencia = EvidenciaActividadDetalleDistributivo(**_dict)
                                        _evidencia.criterio = distributivo

                                    _evidencia.save()

                                    _evidencia.anexoevidenciaactividad_set.filter(status=True).delete()
                                    for anexo in evidencia.anexoevidenciaactividad_set.filter(status=True):
                                        a = AnexoEvidenciaActividad(evidencia=_evidencia, observacion=anexo.observacion,
                                                                    archivo=anexo.archivo)
                                        a.save()

                                    _evidencia.evidenciaactividadaudi_set.filter(status=True).delete()
                                    for anexo in evidencia.evidenciaactividadaudi_set.filter(status=True):
                                        a = EvidenciaActividadAudi(evidencia=_evidencia, archivo=anexo.archivo)
                                        a.save()

                                    _evidencia.historialaprobacionevidenciaactividad_set.filter(status=True).delete()
                                    for anexo in evidencia.historialaprobacionevidenciaactividad_set.filter(
                                            status=True):
                                        model = HistorialAprobacionEvidenciaActividad(evidencia=_evidencia,
                                                                                      aprobacionpersona=anexo.aprobacionpersona,
                                                                                      observacion=anexo.observacion,
                                                                                      fechaaprobacion=anexo.fechaaprobacion,
                                                                                      estadoaprobacion=anexo.estadoaprobacion)
                                        model.save()
                                else:
                                    if DetalleDistributivo.objects.filter(
                                            distributivo__profesor__persona=integrante.persona,
                                            distributivo__periodo=detalledistributivo.distributivo.periodo,
                                            distributivo__activo=True, status=True).values('id').exists():
                                        gnro = "a" if integrante.persona.es_mujer() else "o"
                                        msj = (
                                            f"""Estimad{gnro} {integrante.persona.__str__().lower().title()}, usted se encuentra asociad{gnro} al grupo de investigación "{director.grupo.nombre}" como <b>{integrante.get_funcion_display().lower().title()}</b> pero no cuenta con el criterio<b>{detalledistributivo.criterioinvestigacionperiodo.criterio.nombre.lower().title()}</b> en su distributivo de horas.<br><br> Por favor comuníquese con su director de carrera.""")
                                        notificacion2("Problemas en el distributivo del docente", msj,
                                                      integrante.persona, None, 'notificacion',
                                                      CRITERIO_INTEGRANTE_DIRECTOR_GRUPO_INVESTIGACION, 1, 'sga',
                                                      CriterioInvestigacion)

                        except Exception as ex:
                            print(ex.__str__())
        print("%s" % directory)
    except Exception as ex:
        print(ex.__str__())


def actualizar_fecha_fin_bitacora_diciembre_v2():
    try:
        from inno.models import InformeMensualDocente
        from sga.templatetags.sga_extras import encrypt

        libre_origen = '/actualizacion_bitacora_diciembre_2023.xls'
        fuentecabecera = easyxf(
            'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
        fuentenormal = easyxf(
            'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
        output_folder = MEDIA_ROOT
        output_folder = os.path.join(os.path.join(BASE_DIR))
        # liborigen = xlrd.open_workbook(output_folder + libre_origen)
        libdestino = xlwt.Workbook()
        hojadestino = libdestino.add_sheet('HOJA1')
        fil = 0
        columnas = [(u"CEDULA", 6000), (u"APELLIDOS Y NOMBRES", 6000), (u"CORREO", 6000)]
        for col_num in range(len(columnas)):
            hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
            hojadestino.col(col_num).width = columnas[col_num][1]
        fila = 1
        lin = 0
        miarchivo = openpyxl.load_workbook("media/importar_permisos_loes_diciembre_2023.xlsx")
        lista = miarchivo.get_sheet_by_name('Hoja1')
        a = 0
        for filas in lista.rows:
            if f"{filas[0].value}".isdigit():
                ci = f"{filas[0].value}"
                if persona := Persona.objects.filter(cedula=ci if ci.__len__() == 10 else f"0{ci}").first():
                    if filas[4].value.day > 15:
                        for detalle in DetalleDistributivo.objects.filter(
                                Q(criteriodocenciaperiodo__llenarbitacora=True) | Q(
                                        criterioinvestigacionperiodo__llenarbitacora=True) | Q(
                                        criteriogestionperiodo__llenarbitacora=True)).filter(distributivo__periodo=224,
                                                                                             distributivo__profesor__persona=persona).order_by(
                                'distributivo__profesor__persona__apellido1',
                                'distributivo__profesor__persona__apellido2',
                                'distributivo__profesor__persona__nombres'):
                            if bitacora := detalle.bitacoraactividaddocente_set.filter(fechaini__month=12,
                                                                                       fechaini__year=2023,
                                                                                       profesor__persona=persona,
                                                                                       status=True).first():
                                if not InformeMensualDocente.objects.filter(distributivo=detalle.distributivo,
                                                                            fechafin__month=12, fechafin__year=2023,
                                                                            estado__in=(2, 3, 4)).exists():
                                    a += 1
                                    bitacora.fechafin = filas[4].value
                                    bitacora.save()

                                    cri = ''
                                    if detalle.criteriodocenciaperiodo: cri = detalle.criteriodocenciaperiodo.criterio
                                    if detalle.criterioinvestigacionperiodo: cri = detalle.criterioinvestigacionperiodo.criterio
                                    if detalle.criteriogestionperiodo: cri = detalle.criteriogestionperiodo.criterio

                                    notificacion2("Actualización de la bitácora de actividades",
                                                  f"Estimad{'a' if persona.es_mujer() else 'o'} docente, se informa que se actualizó la fecha fin de la bitácora de diciembre 2023 para la actividad <b>{cri}</b> en función de la fecha de inicio de sus vacaciones. Por favor registrar los {bitacora.fechafin.day - 15} días faltantes.",
                                                  persona, None,
                                                  f'/pro_cronograma?action=detallebitacora&idbitacora={encrypt(bitacora.pk)}',
                                                  1, 1, 'sga', DetalleDistributivo)

                                    # Crea archivo excel
                                    hojadestino.write(fila, 0, "%s" % persona.cedula, fuentenormal)
                                    hojadestino.write(fila, 1, "%s" % persona, fuentenormal)
                                    hojadestino.write(fila, 2, "%s" % persona.emailinst, fuentenormal)

                                    fila += 1
                                    lin += 1
                                    print(
                                        f'{a}. Se actualizó la fecha fin {filas[4].value} en la bitácora del/a docente {persona}')
        libdestino.save(output_folder + libre_origen)
        notificacion2("Proceso finalizado",
                      f"Se generó el archivo con los resultados en la ruta: <a href='{output_folder + libre_origen}'>{libre_origen}</a>",
                      Persona.objects.get(cedula='0606274652'), None, '', 1, 1, 'sga', DetalleDistributivo)
        print(output_folder + libre_origen)
    except Exception as ex:
        pass


def actualizar_fecha_fin_bitacora_diciembre():
    try:
        from sagest.models import PermisoInstitucionalDetalle
        from inno.models import InformeMensualDocente

        BitacoraActividadDocente.objects.filter(fechafin__month=12, fechafin__year=2023, fechafin__day=31,
                                                status=True).update(status=False)

        count = 0
        detalle = DetalleDistributivo.objects.filter(
            Q(criteriodocenciaperiodo__llenarbitacora=True) | Q(criterioinvestigacionperiodo__llenarbitacora=True) | Q(
                criteriogestionperiodo__llenarbitacora=True)).filter(distributivo__periodo=224).order_by(
            'distributivo__profesor__persona__apellido1', 'distributivo__profesor__persona__apellido2',
            'distributivo__profesor__persona__nombres')
        for d in detalle:
            persona = d.distributivo.profesor.persona
            if permiso := PermisoInstitucionalDetalle.objects.filter(permisoinstitucional__regimenlaboral_id=2,
                                                                     fechainicio__month=12, fechainicio__year=2023,
                                                                     status=True, permisoinstitucional__status=True,
                                                                     permisoinstitucional__estadosolicitud__in=(3, 5),
                                                                     permisoinstitucional__tipopermiso_id=24,
                                                                     permisoinstitucional__solicita=persona).first():
                if permiso.fechainicio.day > 15:
                    if bitacora := BitacoraActividadDocente.objects.filter(criterio=d, profesor__persona=persona,
                                                                           status=True).first():
                        if not InformeMensualDocente.objects.filter(distributivo=d.distributivo, fechafin__month=12,
                                                                    fechafin__year=2023, estado__in=(2, 3, 4)).exists():
                            bitacora.fechafin = permiso.fechainicio
                            bitacora.save()
                            count += 1

                            # Envío de correo
                            print(
                                f'{count}. Se actualizó la fecha fin {permiso.fechainicio} en la bitácora del/a docente {persona}')
    except Exception as ex:
        pass


def actualizar_actividad_internado_rotativo():
    from sga.models import Persona, DetalleDistributivo, CriterioDocenciaPeriodo, ClaseActividadEstado
    persona = Persona.objects.get(pk=37121)
    try:
        periodo = 306
        if not CriterioDocenciaPeriodo.objects.filter(periodo=periodo, criterio=167):
            item = CriterioDocenciaPeriodo.objects.filter(periodo=224, criterio=167).values()[0]
            item.pop('id')
            item['periodo_id'] = periodo
            CriterioDocenciaPeriodo(**item).save()
        count = 0
        criteriodocenciaperiodo = CriterioDocenciaPeriodo.objects.filter(periodo=periodo, criterio=167).first()
        listtorollback = []
        for detalledistributivo in DetalleDistributivo.objects.filter(criteriodocenciaperiodo__criterio=133,
                                                                      criteriodocenciaperiodo__periodo=periodo,
                                                                      distributivo__activo=True, status=True,
                                                                      distributivo__status=True)[0:1]:
            distributivo = detalledistributivo.distributivo
            distributivo.bloqueardistributivo = False
            distributivo.save()

            uncheck = False
            # Para que no les notifique el cambio de actividad
            if claseactividadestado := ClaseActividadEstado.objects.filter(profesor=distributivo.profesor,
                                                                           periodo=periodo, estadosolicitud=2,
                                                                           status=True).first():
                claseactividadestado.estadosolicitud = 1
                claseactividadestado.save()
                uncheck = True

            detalledistributivo.criteriodocenciaperiodo = criteriodocenciaperiodo
            detalledistributivo.save()
            detalledistributivo.actualiza_padre()

            if actividaddetalledistributivo := detalledistributivo.actividaddetalledistributivo_set.filter(
                    vigente=True).first():
                actividaddetalledistributivo.nombre = f"{criteriodocenciaperiodo.criterio.nombre}".strip()
                actividaddetalledistributivo.save()

            # Cambiar la actividad planificada en el horario de actividades
            # ClaseActividad.objects.filter(detalledistributivo=detalledistributivo, activo=True, status=True)

            if uncheck:
                claseactividadestado.estadosolicitud = 2
                ClaseActividad.objects.filter(detalledistributivo__distributivo=distributivo, activo=True,
                                              status=True).update(estadosolicitud=2)
                claseactividadestado.save()

            distributivo.bloqueardistributivo = True
            distributivo.save()
            count += 1
            listtorollback.append(detalledistributivo.pk)
        notificacion2(f"Lista para revertir en caso de error {count}", "%s" % listtorollback, persona, None,
                      'notificacion', persona.pk, 1, 'sga', Persona)
    except Exception as ex:
        linea_error = '{}'.format(sys.exc_info()[-1].tb_lineno)
        notificacion2("Problemas en el arreglo", f"ERROR: {ex.__str__()} - {linea_error=}", persona, None,
                      'notificacion', persona.pk, 1, 'sga', Persona)


def actualizar_fecha_actividades():
    from inno.models import InformeMensualDocente
    with transaction.atomic():
        try:
            if DEBUG:
                path_anexo = "media/docentes_y_tecnicos_periodo_224.xlsx"

            miarchivo = openpyxl.load_workbook(path_anexo)
            lista = miarchivo.get_sheet_by_name('OCASIONALES')
            count = 0
            nuevocriteriodocencia = 1187
            subactividad = 'APLICAR EXAMENES PARCIALES O FINALES'
            for filas in lista.rows:
                cedula, periodo, modelexample = f"{filas[0].value}", 224, None
                if cedula.isdigit():
                    porcentaje = float(filas[2].value.replace('%', ''))
                    if not porcentaje == 100:
                        if distributivo := ProfesorDistributivoHoras.objects.filter(profesor__persona__cedula=cedula,
                                                                                    periodo=periodo, activo=True,
                                                                                    status=True).first():
                            if not InformeMensualDocente.objects.filter(distributivo=distributivo, fechafin__year=2024,
                                                                        fechafin__month=1, status=True).values(
                                    'id').exists():
                                distributivo.bloqueardistributivo = False
                                distributivo.save()
                                horas = 0

                                uncheck = False
                                if claseactividadestado := ClaseActividadEstado.objects.filter(
                                        profesor=distributivo.profesor, periodo=periodo, estadosolicitud=2,
                                        status=True).first():
                                    claseactividadestado.estadosolicitud = 1
                                    claseactividadestado.save()
                                    uncheck = True

                                ddistributivo = DetalleDistributivo.objects.filter(distributivo=distributivo,
                                                                                   status=True)
                                pks, iddoc, idinv, idges = [], [], [], []

                                if doc := filas[5].value:
                                    iddoc = f"{doc}".strip().split(',')

                                if inv := filas[6].value:
                                    idinv = f"{inv}".strip().split(',')

                                if ges := filas[7].value:
                                    idges = f"{ges}".strip().split(',')

                                if d := ddistributivo.filter(
                                        Q(criteriodocenciaperiodo__in=iddoc) | Q(criteriogestionperiodo__in=idges) | Q(
                                                criterioinvestigacionperiodo__in=idinv)):
                                    pks += list(d.values_list('id', flat=True))

                                if detallesdistributivodocenciagestion := DetalleDistributivo.objects.filter(pk__in=pks):  # .filter(Q(criteriodocenciaperiodo=1009) | Q(criteriogestionperiodo=1039))
                                    for dd in detallesdistributivodocenciagestion:
                                        if act := dd.actividaddetalledistributivo_set.filter(status=True, vigente=True).first():
                                            act.hasta, act.vigente = date(2023, 12, 31), False
                                            act.save()
                                            act.actualiza_padre()
                                            horas += act.horas

                                    if nuevodetalle := DetalleDistributivo.objects.filter(status=True, distributivo=distributivo, criteriodocenciaperiodo=nuevocriteriodocencia).first():
                                        nuevodetalle.horas = horas
                                    else:
                                        nuevodetalle = DetalleDistributivo(distributivo=distributivo,
                                                                           criteriodocenciaperiodo_id=nuevocriteriodocencia,
                                                                           criterioinvestigacionperiodo=None,
                                                                           criteriogestionperiodo=None,
                                                                           criteriovinculacionperiodo=None,
                                                                           horas=horas,
                                                                           ponderacion_horas=0,
                                                                           es_admision=False)
                                    nuevodetalle.save()

                                    desde, hasta = date(2024, 1, 1), date(2024, 1, 31)
                                    if act := nuevodetalle.actividaddetalledistributivo_set.filter(status=True,
                                                                                                   vigente=True).first():
                                        act.desde = desde
                                        act.hasta = hasta
                                    else:
                                        act = ActividadDetalleDistributivo(criterio=nuevodetalle,
                                                                           nombre=nuevodetalle.criteriodocenciaperiodo.criterio.nombre,
                                                                           desde=desde,
                                                                           hasta=hasta,
                                                                           horas=horas,
                                                                           vigente=True)
                                    act.save()
                                    act.actualiza_padre()

                                    # Cambiar la actividad planificada en el horario de actividades
                                    ClaseActividad.objects.filter(
                                        detalledistributivo__in=detallesdistributivodocenciagestion.values_list('id',
                                                                                                                flat=True),
                                        activo=True, status=True).update(detalledistributivo=nuevodetalle,
                                                                         actividaddetallehorario=act,
                                                                         tipodistributivo=1)

                                    if uncheck:
                                        claseactividadestado.estadosolicitud = 2
                                        ClaseActividad.objects.filter(detalledistributivo__distributivo=distributivo,
                                                                      activo=True, status=True).update(
                                            estadosolicitud=2)
                                        claseactividadestado.save()

                                    distributivo.bloqueardistributivo = True
                                    distributivo.save()
                                    count += 1
                                    print(f"{count}.- {distributivo.profesor.persona}")


        except Exception as ex:
            linea_error = '{}'.format(sys.exc_info()[-1].tb_lineno)
            transaction.set_rollback(True)
            print(f"ERROR: {ex.__str__()} - {linea_error=}")
            persona = Persona.objects.get(cedula='0606274652')
            notificacion2("Problemas en el arreglo", f"ERROR: {ex.__str__()} - {linea_error=}", persona, None,
                          'notificacion', persona.pk, 1, 'sga', Persona)


def generar_informe_cumplimiento():
    from sga.templatetags.sga_extras import listado_bitacora_docente
    try:
        fechaini, fechafin = date(2024, 1, 1), date(2024, 1, 31)
        periodo = 224
        listadoprofesormateria = ProfesorDistributivoHoras.objects.filter(carrera_id__isnull=False, status=True,
                                                                          periodo=periodo, activo=True).order_by(
            'profesor__persona__apellido1', 'profesor__persona__apellido2', 'profesor__persona__nombres')
        resumen = []
        criterios = (
            (1, 'DOCENCIA'),
            (2, 'INVESTIGACIÓN'),
            (3, 'GESTIÓN'),
            (4, 'VINCULACIÓN'),
            (5, 'TOTAL'),
        )
        for distributivo in listadoprofesormateria[10:15]:
            periodo = distributivo.periodo
            profesor = distributivo.profesor
            count, count1, count2, count3, count4 = 0, 0, 0, 0, 0
            totalporcentaje, totalhdocentes, totalhinvestigacion, totalhgestion, totalhvinculacion = 0, 0, 0, 0, 0
            fechames = datetime.now().date()
            now = datetime.now()
            yearini = now.year
            year = now.year
            dayini = 1
            dia = int(now.day)
            fini, ffin = fechaini, fechafin
            fechainiresta = fini - timedelta(days=5)
            fechafinresta = ffin - timedelta(days=5)
            finicresta = fechainiresta
            ffincresta = fechafinresta

            print(f"Calculando: {profesor} {fini} - {ffin}")

            finiinicio = fechaini
            ffinal = fechafin

            adicional_lista = []
            lista_criterios = []
            listDocencia = []
            # print(f'--------------DOCENTES--------------')
            __doc, __inv, __ges, __vin = [0, 0, 0], [0, 0, 0], [0, 0, 0], [0, 0, 0]
            if horasdocencia := distributivo.detalle_horas_docencia(finiinicio, ffinal):
                dicDocencia = {'tipo': 'Horas Docencia'}
                # listDocencia = []
                listDocencia.append([0, 'ACTIVIDADES DE DOCENCIA'])
                for actividad in horasdocencia:
                    if actividad.criteriodocenciaperiodo.nombrehtmldocente():
                        __doc[0] += actividad.horas
                        if actividad.criteriodocenciaperiodo.nombrehtmldocente() == 'impartirclase':
                            profesormateria = ProfesorMateria.objects.filter(profesor=distributivo.profesor,
                                                                             materia__nivel__periodo=periodo,
                                                                             tipoprofesor__imparteclase=True,
                                                                             activo=True,
                                                                             materia__status=True).distinct().order_by(
                                'desde', 'materia__asignatura__nombre')
                            if periodo.clasificacion == 1:
                                asignaturas = profesormateria.filter(((Q(desde__gte=finiinicio) & Q(
                                    hasta__lte=ffinal)) | (Q(desde__lte=finiinicio) & Q(hasta__gte=ffinal)) | (
                                                                                  Q(desde__lte=finiinicio) & Q(
                                                                              desde__gte=ffinal)) | (
                                                                                  Q(hasta__gte=finiinicio) & Q(
                                                                              hasta__lte=ffinal)))).distinct().exclude(
                                    tipoprofesor_id__in=[15, 5]).order_by('desde', 'materia__asignatura__nombre')
                            else:
                                asignaturas = profesormateria.filter(((Q(desde__gte=finiinicio) & Q(
                                    hasta__lte=ffinal)) | (Q(desde__lte=finiinicio) & Q(hasta__gte=ffinal)) | (
                                                                                  Q(desde__lte=finiinicio) & Q(
                                                                              desde__gte=ffinal)) | (
                                                                                  Q(hasta__gte=finiinicio) & Q(
                                                                              hasta__lte=ffinal)))).exclude(
                                    tipoprofesor_id__in=[5]).distinct().order_by('desde', 'materia__asignatura__nombre')
                            totalimpartir = actividad.criteriodocenciaperiodo.totalimparticlase(distributivo.profesor,
                                                                                                finiinicio, ffinal,
                                                                                                asignaturas, None, True)
                            if totalimpartir[2]:
                                count += 1
                                totalhdocentes += totalimpartir[1]
                            __doc[1] += totalimpartir[0]
                            __doc[2] += totalimpartir[1]

                        if actividad.criteriodocenciaperiodo.nombrehtmldocente() == 'evidenciamoodle':
                            if not DEBUG:
                                if listadoevidencias := actividad.criteriodocenciaperiodo.horario_evidencia_moodle(
                                        distributivo.profesor, finicresta, ffincresta, True):
                                    if listadoevidencias[2]:
                                        count += 1
                                        totalhdocentes += listadoevidencias[1]
                                    # listDocencia.append([actividad.criteriodocenciaperiodo.id, actividad.criteriodocenciaperiodo.criterio.nombre, actividad.horas, listadoevidencias[0], listadoevidencias[1]])
                                    __doc[1] += listadoevidencias[0]
                                    __doc[2] += listadoevidencias[1]

                        if actividad.criteriodocenciaperiodo.nombrehtmldocente() == 'materialsilabo':
                            actividadhor = actividad.criteriodocenciaperiodo.horarios_actividad_profesor(
                                distributivo.profesor, finiinicio, ffinal, True)
                            count += 1
                            totalhdocentes += actividadhor[1]
                            __doc[1] += actividadhor[0]
                            __doc[2] += actividadhor[1]
                            # listDocencia.append([actividad.criteriodocenciaperiodo.id, actividad.criteriodocenciaperiodo.criterio.nombre, actividad.horas, actividadhor[0], actividadhor[1]])

                        if actividad.criteriodocenciaperiodo.nombrehtmldocente() == 'cursonivelacion':
                            actividadnivelacioncarrera = actividad.criteriodocenciaperiodo.horarios_nivelacioncarrera_profesor(
                                distributivo.profesor, finiinicio, ffinal)
                            totitem4 = 0
                            if actividadnivelacioncarrera:
                                totitem4 += 100
                                totalhdocentes += 100
                                count += 1
                                __doc[1] += actividad.horas
                                __doc[2] += 100

                        if actividad.criteriodocenciaperiodo.nombrehtmldocente() == 'planificarcontenido':
                            contenidohor = actividad.criteriodocenciaperiodo.horarios_contenido_profesor(
                                distributivo.profesor, finiinicio, ffinal, True)
                            if contenidohor == 0:
                                listDocencia.append([actividad.criteriodocenciaperiodo.id,
                                                     actividad.criteriodocenciaperiodo.criterio.nombre, actividad.horas,
                                                     '-', '-'])
                            else:
                                count += 1
                                totalhdocentes += contenidohor[1]
                                __doc[1] += contenidohor[0]
                                __doc[2] += contenidohor[1]
                                # listDocencia.append([actividad.criteriodocenciaperiodo.id, actividad.criteriodocenciaperiodo.criterio.nombre, actividad.horas, contenidohor[0], contenidohor[1]])

                        if actividad.criteriodocenciaperiodo.nombrehtmldocente() == 'tutoriaacademica':
                            tutoriasacademicas = actividad.criteriodocenciaperiodo.horarios_tutoriasacademicas_profesor(
                                distributivo.profesor, finiinicio, ffinal, True)
                            count += 1
                            totalhdocentes += tutoriasacademicas[1]
                            __doc[1] += tutoriasacademicas[0]
                            __doc[2] += tutoriasacademicas[1]
                            # listDocencia.append([actividad.criteriodocenciaperiodo.id, actividad.criteriodocenciaperiodo.criterio.nombre, actividad.horas, tutoriasacademicas[0], tutoriasacademicas[1]])

                        if actividad.criteriodocenciaperiodo.nombrehtmldocente() == 'seguimientoplataforma':
                            listadoseguimientos = actividad.criteriodocenciaperiodo.horario_seguimiento_tutor_fecha(
                                distributivo.profesor, finiinicio, ffinal, True)
                            if listadoseguimientos[2]:
                                count += 1
                                totalhdocentes += listadoseguimientos[1]
                            __doc[1] += listadoseguimientos[0]
                            __doc[2] += listadoseguimientos[1]
                            # listDocencia.append([actividad.criteriodocenciaperiodo.id, actividad.criteriodocenciaperiodo.criterio.nombre, actividad.horas, listadoseguimientos[0], listadoseguimientos[1]])

                        if actividad.criteriodocenciaperiodo.nombrehtmldocente() == 'nivelacioncarrera':
                            actividadgestion = actividad.criteriodocenciaperiodo.horarios_informesdocencia_profesor(
                                distributivo, finiinicio, ffinal, True)
                            count += 1
                            totalhdocentes += actividadgestion[1]
                            __doc[1] += actividadgestion[0]
                            __doc[2] += actividadgestion[1]
                            # listDocencia.append([actividad.criteriodocenciaperiodo.id, actividad.criteriodocenciaperiodo.criterio.nombre, actividad.horas, actividadgestion[0], actividadgestion[1]])

                        if actividad.criteriodocenciaperiodo.nombrehtmldocente() == 'seguimientotransversal':
                            listadoseguimientos = actividad.criteriodocenciaperiodo.horario_seguimiento_transaversal_fecha(
                                distributivo.profesor, finiinicio, ffinal, True)
                            if listadoseguimientos[2]:
                                count += 1
                                totalhdocentes += listadoseguimientos[1]
                            __doc[1] += listadoseguimientos[0]
                            __doc[2] += listadoseguimientos[1]
                            # listDocencia.append([actividad.criteriodocenciaperiodo.id, actividad.criteriodocenciaperiodo.criterio.nombre, actividad.horas, listadoseguimientos[0], listadoseguimientos[1]])

                        if actividad.criteriodocenciaperiodo.nombrehtmldocente() == 'apoyovicerrectorado':
                            actividadapoyo = actividad.criteriodocenciaperiodo.horarios_apoyo_profesor(
                                distributivo.profesor, finiinicio, ffinal)
                            totitem10 = 0
                            if actividadapoyo:
                                totitem10 += 100
                                totalhdocentes += 100
                                count += 1
                                # __doc[1] += claseactividad['totalplanificadas']
                                __doc[2] += 100

                        if actividad.criteriodocenciaperiodo.nombrehtmldocente() == 'actividaddocente':
                            actividaddocente1 = actividad.criteriodocenciaperiodo.horarios_actividaddocente_profesor(
                                distributivo.profesor, finiinicio, ffinal, True)
                            if actividaddocente1:
                                count += 1
                                totalhdocentes += 100
                                __doc[1] += actividaddocente1[0]
                                __doc[2] += actividaddocente1[1]
                            else:
                                count += 1
                                totalhdocentes += 0
                                __doc[2] += 0

                        if actividad.criteriodocenciaperiodo.nombrehtmldocente() == 'criterioperiodoadmision':
                            actividaddocente1 = actividad.criteriodocenciaperiodo.horario_criterio_nivelacion(
                                distributivo.profesor, finiinicio, ffinal, True)
                            count += 1
                            totalhdocentes += actividaddocente1[1]
                            __doc[1] += actividaddocente1[0]
                            __doc[2] += actividaddocente1[1]
                            # listDocencia.append([actividad.criteriodocenciaperiodo.id,actividad.criteriodocenciaperiodo.criterio.nombre, actividad.horas, actividaddocente1[0], actividaddocente1[1]])

                        if actividad.criteriodocenciaperiodo.nombrehtmldocente() == 'actividadbitacora':
                            actividaddocente1 = listado_bitacora_docente(0, actividad, ffinal, True)
                            count += 1
                            totalhdocentes += actividaddocente1[1]
                            __doc[1] += actividaddocente1[0]
                            __doc[2] += actividaddocente1[1]
                            # listDocencia.append([actividad.criteriodocenciaperiodo.id, actividad.criteriodocenciaperiodo.criterio.nombre, actividad.horas, actividaddocente1[0], actividaddocente1[1]])

                adicional_lista.append(listDocencia)
                lista_criterios.append({'tipocriterio': 1, 'data': listDocencia})
                listDocencia = []

            if horasinvestigacion := distributivo.detalle_horas_investigacion():
                docInvestigacion = {'tipo': 'Horas Investigación'}
                listDocencia.append([0, 'ACTIVIDADES DE INVESTIGACIÓN'])
                listInvestigacion = []
                for actividad in horasinvestigacion:
                    if actividad.criterioinvestigacionperiodo.nombrehtmldocente():
                        __inv[0] += actividad.horas
                        if actividad.criterioinvestigacionperiodo.nombrehtmldocente() == 'actividadinvestigacion':
                            actividadgestion = actividad.criterioinvestigacionperiodo.horarios_informesinvestigacion_profesor(
                                distributivo, finiinicio, ffinal, True)
                            count1 += 1
                            totalhinvestigacion += actividadgestion[1]
                            __inv[1] += actividadgestion[0]
                            __inv[2] += actividadgestion[1]
                            # listDocencia.append([actividad.criterioinvestigacionperiodo.id, actividad.criterioinvestigacionperiodo.criterio.nombre, actividad.horas, '', actividadgestion[1]])

                        if actividad.criterioinvestigacionperiodo.nombrehtmldocente() == 'actividadbitacora':
                            actividadgestion = listado_bitacora_docente(0, actividad, ffinal, True)
                            count1 += 1
                            totalhinvestigacion += actividadgestion[1]
                            __inv[1] += actividadgestion[0]
                            __inv[2] += actividadgestion[1]
                            # listDocencia.append([actividad.criterioinvestigacionperiodo.id, actividad.criterioinvestigacionperiodo.criterio.nombre, actividad.horas, hmes, actividadgestion[1]])
                lista_criterios.append({'tipocriterio': 2, 'data': listDocencia})
                listDocencia = []

            if horasgestion := distributivo.detalle_horas_gestion(finiinicio, ffinal):
                docGestion = {'tipo': 'Horas Gestión'}
                listGestion = []
                listDocencia.append([0, 'ACTIVIDADES DE GESTIÓN EDUCATIVA'])
                for actividad in horasgestion:
                    if actividad.criteriogestionperiodo.nombrehtmldocente():
                        __ges[0] += actividad.horas
                        if actividad.criteriogestionperiodo.nombrehtmldocente() == 'actividadgestion':
                            actividadgestion = actividad.criteriogestionperiodo.horarios_actividadgestion_profesor(
                                distributivo.profesor, finiinicio, ffinal, True)
                            if actividadgestion:
                                count2 += 1
                                totalhgestion += 100
                                __ges[1] += actividadgestion[0]
                                __ges[2] += float(actividadgestion[1])
                                # listDocencia.append([actividad.criteriogestionperiodo.id, actividad.criteriogestionperiodo.criterio.nombre, actividad.horas, actividadgestion[0], actividadgestion[1]])
                            else:
                                count2 += 1
                                totalhgestion += 0
                                __ges[1] += 0
                                __ges[2] += 0
                                # listDocencia.append([actividad.criteriogestionperiodo.id, actividad.criteriogestionperiodo.criterio.nombre, actividad.horas, '-', '0.00'])
                        if actividad.criteriogestionperiodo.nombrehtmldocente() == 'actividadinformegestion':
                            actividadgestion = actividad.criteriogestionperiodo.horarios_informesgestion_profesor(
                                distributivo, finiinicio, ffinal, True)
                            count2 += 1
                            totalhgestion += actividadgestion[1]
                            __ges[1] += actividadgestion[0]
                            __ges[2] += float(actividadgestion[1])
                            # listDocencia.append([actividad.criteriogestionperiodo.id, actividad.criteriogestionperiodo.criterio.nombre,actividad.horas, '', actividadgestion[1]])

                        if actividad.criteriogestionperiodo.nombrehtmldocente() == 'actividadbitacora':
                            actividadgestion = listado_bitacora_docente(0, actividad, ffinal, True)
                            count2 += 1
                            totalhgestion += actividadgestion[1]
                            __ges[1] += actividadgestion[0]
                            __ges[2] += float(actividadgestion[1])
                            # listDocencia.append([actividad.criteriogestionperiodo.id, actividad.criteriogestionperiodo.criterio.nombre,actividad.horas, actividadgestion[0], actividadgestion[1]])
                lista_criterios.append({'tipocriterio': 3, 'data': listDocencia})
                listDocencia = []

            if horasvinculacion := distributivo.detalle_horas_vinculacion():
                docVinculacion = {'tipo': 'Horas Vinculacion'}
                listVinculacion = []
                listDocencia.append([0, 'ACTIVIDADES DE VINCULACIÓN CON LA SOCIEDAD'])
                for actividad in horasvinculacion:
                    if actividad.criteriodocenciaperiodo.nombrehtmldocente():
                        __vin[0] += actividad.horas
                        if actividad.criteriodocenciaperiodo.nombrehtmldocente() == 'actividadvinculacion':
                            actividadgestion = actividad.criteriodocenciaperiodo.horarios_informesdocencia_profesor(
                                distributivo, finiinicio, ffinal, True)
                            if actividadgestion:
                                count3 += 1
                                totalhvinculacion += actividadgestion[1]
                                __vin[1] += actividadgestion[0]
                                __vin[2] += actividadgestion[1]
                                # listDocencia.append([actividad.criteriodocenciaperiodo.id, actividad.criteriodocenciaperiodo.criterio.nombre, actividad.horas,actividadgestion[0], actividadgestion[1]])
                            else:
                                count3 += 1
                                totalhvinculacion += 0
                                __vin[1] += 0
                                __vin[2] += 0
                                # listDocencia.append([actividad.criteriodocenciaperiodo.id,actividad.criteriodocenciaperiodo.criterio.nombre, actividad.horas,'-', '0.00'])

                        if actividad.criteriodocenciaperiodo.nombrehtmldocente() == 'actividadbitacora':
                            actividadgestion = listado_bitacora_docente(0, actividad, ffinal, True)
                            count3 += 1
                            totalhvinculacion += actividadgestion[1]
                            __vin[1] += actividadgestion[0]
                            __vin[2] += actividadgestion[1]
                            # listDocencia.append([actividad.criteriodocenciaperiodo.id,actividad.criteriodocenciaperiodo.criterio.nombre, actividad.horas, actividadgestion[0], actividadgestion[1]])

                lista_criterios.append({'tipocriterio': 4, 'data': listDocencia})
                listDocencia = []

            totalporcentaje = totalhdocentes + totalhinvestigacion + totalhgestion + totalhvinculacion
            count4 = count + count1 + count2 + count3
            total_porcentaje = round(totalporcentaje / count4 if count4 else totalporcentaje, 2)
            listDocencia.append(['total', total_porcentaje])
            lista_criterios.append({'tipocriterio': 5, 'data': total_porcentaje})
            resumen.append({'profesor': profesor, 'data': lista_criterios})
        print(resumen.__str__())
    except Exception as ex:
        pass


def migrar_evidencia_vinculacion():
    try:
        from inno.models import MigracionEvidenciaActividad
        from sga.templatetags.sga_extras import nombremes
        from core.firmar_documentos import verificarFirmasPDF
        from sga.funciones import notificacion2
        data_set = []
        # -- la prueba es de alguien que no se le migró y otro que tiene marzo
        lista_prueba = [21700, 36]
        for configuracion in ConfiguracionInformeVinculacion.objects.filter(profesor__persona__in=lista_prueba,
                                                                            fecha_fin__month__in=(2, 3),
                                                                            fecha_fin__year=2024, aprobacion__gte=2,
                                                                            status=True):
            owner = configuracion.profesor.persona
            lider = configuracion.proyecto.lider()
            pk_criterio = [151, 150][owner == lider]
            c = CriterioDocencia.objects.get(pk=pk_criterio)
            periodo = Periodo.objects.get(pk=306)
            if criterio := DetalleDistributivo.objects.filter(criteriodocenciaperiodo__criterio=c,
                                                              criteriodocenciaperiodo__periodo=periodo,
                                                              distributivo__profesor__persona=owner,
                                                              distributivo__periodo_id=periodo, status=True).first():
                evidencia = None

                if migracion := MigracionEvidenciaActividad.objects.filter(informevinculacion=configuracion,
                                                                           status=True).first():
                    evidencia = migracion.evidencia

                if temporal := EvidenciaActividadDetalleDistributivo.objects.filter(
                        proyectovinculacion=configuracion.proyecto, criterio=criterio, desde=configuracion.fecha_inicio,
                        hasta=configuracion.fecha_fin, generado=True).first():
                    if not evidencia:
                        evidencia = temporal

                valido, _, dict = verificarFirmasPDF(configuracion.archivo)
                num_firmas = dict.get('firmasValidas').__len__() if dict else 0

                if not evidencia:
                    estadoaprobacion = 1
                    if num_firmas == 2:
                        estadoaprobacion = 4
                        configuracion.aprobacion = 3

                    if num_firmas == 3:
                        estadoaprobacion = 5
                        configuracion.aprobacion = 3

                    if configuracion.aprobacion == 4:
                        estadoaprobacion = 3

                    evidencia = EvidenciaActividadDetalleDistributivo(estadoaprobacion=estadoaprobacion,
                                                                      archivo=configuracion.archivo,
                                                                      archivofirmado=configuracion.archivo,
                                                                      proyectovinculacion=configuracion.proyecto,
                                                                      criterio=criterio,
                                                                      desde=configuracion.fecha_inicio,
                                                                      hasta=configuracion.fecha_fin,
                                                                      actividad=f'INFORMAR SOBRE LAS ACTIVIDADES REALIZADAS EN EL MES DE {nombremes(configuracion.fecha_fin).upper()}',
                                                                      generado=True)
                    evidencia.save()
                    historial = HistorialAprobacionEvidenciaActividad(evidencia=evidencia,
                                                                      observacion=u'Cumple con lo solicitado',
                                                                      estadoaprobacion=estadoaprobacion,
                                                                      aprobacionpersona=lider,
                                                                      fechaaprobacion=datetime.now().date())
                    historial.save()
                    if not migracion:
                        migracion = MigracionEvidenciaActividad(evidencia=evidencia, informevinculacion=configuracion)
                        migracion.save()
                else:
                    if num_firmas >= 2 and not evidencia.estadoaprobacion == 5:
                        evidencia.archivo = configuracion.archivo
                        evidencia.estadoaprobacion = 4
                        evidencia.save()

                    archivo = evidencia.archivofirmado if evidencia.archivofirmado else evidencia.archivo
                    valido, _, dict = verificarFirmasPDF(archivo)
                    num_firmas_ev = dict.get('firmasValidas').__len__() if dict else 0

                    if num_firmas_ev >= 2 and num_firmas < 2:
                        configuracion.archivo = evidencia.archivofirmado
                        configuracion.aprobacion = 3
                        configuracion.save()

                    if evidencia.estadoaprobacion == 3:
                        configuracion.aprobacion = 4
                    if evidencia.estadoaprobacion in (2, 4, 5):
                        configuracion.aprobacion = 3

                    configuracion.save()
                data_set.append(evidencia.pk)
            DEBUG and print("%s" % configuracion.profesor.persona)

        persona = Persona.objects.get(cedula='0606274652')
        notificacion2("Resultados arreglo migrar_evidencia_vinculacion", f"{data_set=}", persona, None, '/notificacion',
                      persona.pk, 1, 'sga', Persona)
        DEBUG and print(data_set)
    except Exception as ex:
        linea_error = 'Error on line {}'.format(sys.exc_info()[-1].tb_lineno)
        persona = Persona.objects.get(cedula='0606274652')
        DEBUG and print(f"{ex=}. {linea_error}")
        notificacion2("Resultados", f"{ex=}. {linea_error}", persona, None, '/notificacion', persona.pk, 1, 'sga',
                      Persona)


def migracion_actividades_periodo_remedial():
    try:
        from sga.funciones import notificacion2
        periodo = Periodo.objects.get(pk=306)
        __response = []
        for dd in DetalleDistributivo.objects.filter(Q(distributivo__periodo=periodo, status=True)):
            distributivo = dd.distributivo
            distributivo.bloqueardistributivo = False
            distributivo.save()

            uncheck = False
            if claseactividadestado := ClaseActividadEstado.objects.filter(profesor=distributivo.profesor,
                                                                           periodo=periodo, estadosolicitud=2,
                                                                           status=True).first():
                claseactividadestado.estadosolicitud = 1
                claseactividadestado.save()
                uncheck = True

            if doc := dd.criteriodocenciaperiodo:
                if not doc.periodo == periodo:
                    if doc_n := CriterioDocenciaPeriodo.objects.filter(criterio=doc.criterio, periodo=periodo,
                                                                       status=True).first():
                        dd.criteriodocenciaperiodo = doc_n
                        dd.save()
                        __response.append(dd.pk)

            if inv := dd.criterioinvestigacionperiodo:
                if not inv.periodo == periodo:
                    if inv_n := CriterioInvestigacionPeriodo.objects.filter(criterio=inv.criterio, periodo=periodo,
                                                                            status=True).first():
                        dd.criterioinvestigacionperiodo = inv_n
                        dd.save()
                        __response.append(dd.pk)

            if ges := dd.criteriogestionperiodo:
                if not ges.periodo == periodo:
                    if ges_n := CriterioGestionPeriodo.objects.filter(criterio=ges.criterio, periodo=periodo,
                                                                      status=True).first():
                        dd.criteriogestionperiodo = ges_n
                        dd.save()
                        __response.append(dd.pk)

            if uncheck:
                claseactividadestado.estadosolicitud = 2
                claseactividadestado.save()

        persona = Persona.objects.get(cedula='0606274652')
        cuerpo = '%s' % DetalleDistributivo.objects.filter(id__in=__response).values(
            'distributivo__profesor__persona__cedula').query.__str__()
        notificacion2("Resultados arreglo migracion_actividades_periodo_remedial", cuerpo, persona, None,
                      '/notificacion', persona.pk, 1, 'sga', Persona)
        DEBUG and print(__response)
    except Exception as ex:
        pass


def migrar_actividad_macro_investigacion():
    from sga.funciones import notificacion2
    persona = Persona.objects.get(id=37121)
    try:
        from inno.models import SubactividadDocentePeriodo, SubactividadDetalleDistributivo
        periodo = Periodo.objects.get(id=336)
        response_str = 'Listado de docentes afectados: '
        for i, detalle in enumerate(DetalleDistributivo.objects.filter(distributivo__periodo=periodo, criterioinvestigacionperiodo__criterio__id=68, status=True)):
            if actividad := detalle.actividaddetalledistributivo_set.filter(status=True, vigente=True).first():
                for subactividad in SubactividadDocentePeriodo.objects.filter(actividad__criterioinvestigacionperiodo=detalle.criterioinvestigacionperiodo, actividad__status=True, status=True):
                    if not SubactividadDetalleDistributivo.objects.values('id').filter(actividaddetalledistributivo=actividad, subactividaddocenteperiodo=subactividad,status=True).exists():
                        s = SubactividadDetalleDistributivo(actividaddetalledistributivo=actividad, subactividaddocenteperiodo=subactividad, fechainicio=actividad.desde, fechafin=actividad.hasta)
                        s.save()
            response_str += f"""
                <br>{i}, {detalle.pk}, {detalle.distributivo.profesor.persona.__str__()}
            """
        notificacion2("Resultados arreglo jefferson", response_str, persona, None, '/notificacion', persona.pk, 1, 'sga', Persona)
        DEBUG and print(response_str)
    except Exception as ex:
        linea_error = 'Error on line {}'.format(sys.exc_info()[-1].tb_lineno)
        notificacion2("Resultados arreglo jefferson", f"{linea_error}. {ex.__str__()}", persona, None, '/notificacion', persona.pk, 1, 'sga', Persona)
#funcionlocal.get("migrar_actividad_macro_investigacion")()

def migrar_actividad_macro_vinculacion():
    from sga.funciones import notificacion2
    from sga.models import Persona, DetalleDistributivo, ActividadDetalleDistributivo
    from inno.models import SubactividadDocentePeriodo, SubactividadDetalleDistributivo, Criterio
    persona = Persona.objects.get(id=37121)
    try:
        periodo = Periodo.objects.get(id=317)
        ACTIVIDAD_MACRO_VIN = 184
        response_str = 'Listado de docentes afectados: '
        no_migrados = ''
        for i, detalle in enumerate(DetalleDistributivo.objects.filter(distributivo__periodo=periodo,
                                                                       criteriodocenciaperiodo__criterio__id=ACTIVIDAD_MACRO_VIN,
                                                                       status=True)):
            distributivo = detalle.distributivo
            distributivo.bloqueardistributivo = False
            horas = sum(list(detalle.actividades().values_list('horas', flat=True)))
            # Se crea la cabecera de la tabla SubactividadDetalleDistributivo
            actividad = ActividadDetalleDistributivo.objects.filter(criterio=detalle,
                                                                    nombre__icontains=f"{detalle.criteriodocenciaperiodo.criterio}".strip(),
                                                                    status=True).first()
            if not actividad:
                actividad = ActividadDetalleDistributivo(criterio=detalle,
                                                         nombre=f"{detalle.criteriodocenciaperiodo.criterio}".strip(),
                                                         desde=periodo.inicio, hasta=periodo.fin, vigente=True,
                                                         horas=horas)
                actividad.save()

            for act in detalle.actividaddetalledistributivo_set.filter(status=True).exclude(id=actividad.pk):
                # Modificar la linea sgte si los datos se leen desde un .xlsx
                if criterio := Criterio.objects.filter(nombre__icontains=f"{act.nombre}".upper().replace('.', '')):
                    # Si la actividad se parece al criterio
                    if criterio.__len__() == 1:
                        subactividaddocenteperiodo = SubactividadDocentePeriodo.objects.filter(
                            actividad__criteriodocenciaperiodo=detalle.criteriodocenciaperiodo, criterio=criterio[0],
                            status=True).first()
                        if not SubactividadDetalleDistributivo.objects.values('id').filter(
                                actividaddetalledistributivo=actividad,
                                subactividaddocenteperiodo=subactividaddocenteperiodo, status=True).exists():
                            subactividad = SubactividadDetalleDistributivo(actividaddetalledistributivo=actividad,
                                                                           subactividaddocenteperiodo=subactividaddocenteperiodo,
                                                                           fechafin=periodo.inicio,
                                                                           fechainicio=periodo.fin)
                            subactividad.save()
                    else:
                        no_migrados += f'{detalle.distributivo.profesor.persona.cedula}, '

                    # Esta funcionalidad está en la accion de eliminar actividad 'delactividad'
                    criterio = act.criterio
                    act.delete()
                    criterio.distributivo.resumen_evaluacion_acreditacion().actualizar_resumen()
            distributivo.bloqueardistributivo = True

            detalle.actualiza_padre()
            response_str += f"{i}, {detalle.pk}, {detalle.distributivo.profesor.persona.__str__()} <br>"
            DEBUG and print(f"{i}, {detalle.pk}, {detalle.distributivo.profesor.persona.__str__()}")
        notificacion2("Resultados arreglo jefferson", response_str + f'<br><br>No migrados: {no_migrados}', persona,
                      None, '/notificacion', persona.pk, 1, 'sga', Persona)
    except Exception as ex:
        linea_error = 'Error on line {}'.format(sys.exc_info()[-1].tb_lineno)
        notificacion2("Resultados arreglo jefferson", f"{linea_error}. {ex.__str__()}", persona, None, '/notificacion',
                      persona.pk, 1, 'sga', Persona)


def generar_criterios_vinculacion():
    from sga.models import Persona, Periodo
    persona = Persona.objects.get(id=37121)
    try:
        from inno.models import ActividadDocentePeriodo, Criterio, SubactividadDocentePeriodo
        from sga.funciones import notificacion2
        periodo = Periodo.objects.get(id=317)
        if not ActividadDocentePeriodo.objects.values('id').filter(criteriodocenciaperiodo_id=1293).exists():
            _actividad = Criterio(
                nombre=u"PRESTAR SERVICIOS A LA SOCIEDAD QUE NO GENEREN BENEFICIO ECONÓMICO PARA LA UNIVERSIDAD O ESCUELA POLITÉCNICA O PARA SU PERSONAL ACADÉMICO, TALES COMO EL ANÁLISIS DE LABORATORIO ESPECIALIZADO, EL PERITAJE JUDICIAL, LA REVISIÓN TÉCNICA DOCUMENTAL PARA LAS INSTITUCIONES DEL ESTADO, ENTRE OTRAS. LA PARTICIPACIÓN REMUNERADA EN TRABAJOS DE CONSULTORÍA INSTITUCIONAL NO SE RECONOCERÁ COMO ACTIVIDAD DE VINCULACIÓN DENTRO DE LA DEDICACIÓN HORARIA",
                tipo=1, tipocriterio=4)
            _actividad.save()

            director = Criterio(nombre=u"DIRECTOR DE PROYECTOS DE VINCULACIÓN CON LA SOCIEDAD", tipo=2, tipocriterio=4)
            director.save()
            asociado = Criterio(nombre=u"ASOCIADO DE PROYECTOS DE VINCULACIÓN CON LA SOCIEDAD", tipo=2, tipocriterio=4)
            asociado.save()

            Criterio(nombre=u"FORMULACIÓN DE PROYECTOS DE SERVICIO COMUNITARIO", tipo=2, tipocriterio=4).save()
            Criterio(
                nombre=u"DIRECCIÓN, TUTORÍAS, SEGUIMIENTO Y EVALUACIÓN DE ACTIVIDADES EXTRACURRICULARES – VINCULACION",
                tipo=2, tipocriterio=4).save()

            actividad = ActividadDocentePeriodo(criterio=_actividad, criteriodocenciaperiodo_id=1293)
            actividad.save()

            for criterio in Criterio.objects.filter(tipo=2, tipocriterio=4, status=True):
                s = SubactividadDocentePeriodo(criterio=criterio,
                                               actividad=actividad,
                                               fechainicio=periodo.inicio,
                                               fechafin=periodo.fin,
                                               validacion=True,
                                               nombrehtml='subactividadinforme',
                                               tipoevidencia=1,
                                               cargaevidencia=False if criterio.pk in [director.pk,
                                                                                       asociado.pk] else True)
                s.save()
            _response = ''
            for r in Criterio.objects.filter(tipocriterio=4, status=True).values_list('id', 'nombre').order_by('id'):
                _response += f'{r} <br>'
                DEBUG and print(r)
            notificacion2("Resultados arreglo jefferson", f"{_response}", persona, None, '/notificacion', persona.pk, 1,
                          'sga', Persona)
    except Exception as ex:
        DEBUG and print(ex.__str__())
        linea_error = 'Error on line {}'.format(sys.exc_info()[-1].tb_lineno)
        notificacion2("Resultados arreglo jefferson", f"{linea_error}. {ex.__str__()}", persona, None, '/notificacion',
                      persona.pk, 1, 'sga', Persona)


def migrar_par_revisor():
    try:
        from sga.models import ParRevisorProduccionCientifica
        from sagest.th_hojavida import migrar_par_revisor
        periodo = Periodo.objects.get(pk=317)
        for parrevisor in ParRevisorProduccionCientifica.objects.filter(
                fecharevision__range=[periodo.inicio, periodo.fin], status=True).order_by('-fecharevision'):
            migrar_par_revisor(parrevisor, persona=parrevisor.persona, periodo=periodo)
            DEBUG and print(f"{parrevisor.persona}")
    except Exception as ex:
        pass


def registrar_horas_vinculacion():
    from sga.funciones import variable_valor
    periodo = Periodo.objects.get(id=317)
    persona = Persona.objects.get(id=37121)
    try:
        response_str = ''
        for detalledistributivo in DetalleDistributivo.objects.filter(distributivo__periodo=periodo,
                                                                      criteriodocenciaperiodo__criterio=variable_valor(
                                                                              'ACTIVIDAD_MACRO_VINCULACION'),
                                                                      criteriodocenciaperiodo__criterio__tipo=2,
                                                                      status=True):
            counter = 0

            uncheck = False
            if claseactividadestado := ClaseActividadEstado.objects.filter(
                    profesor=detalledistributivo.distributivo.profesor, periodo=periodo, estadosolicitud=2,
                    status=True).first():
                claseactividadestado.estadosolicitud = 1
                claseactividadestado.save()
                uncheck = True  # Activa la bandera si el docente ya tenía su horario aprobado

            if actividaddetalle := detalledistributivo.actividaddetalledistributivo_set.filter(vigente=True,
                                                                                               status=True).first():
                if ClaseActividad.objects.filter(actividaddetallehorario=actividaddetalle,
                                                 detalledistributivo=detalledistributivo, activo=True,
                                                 status=True).count() < actividaddetalle.horas:
                    print(f'{detalledistributivo.distributivo.profesor.persona}')
                    for dia in range(1, 6):
                        for turno in Turno.objects.filter(sesion=20, status=True):
                            claseactividad = detalledistributivo.claseactividad_set.filter(turno=turno, dia=dia,
                                                                                           status=True).values(
                                'id').exists()
                            clase = turno.clase_set.filter(dia=dia, profesor=detalledistributivo.distributivo.profesor,
                                                           status=True).values('id').exists()
                            actividades = turno.horario_profesor_actividad(dia,
                                                                           detalledistributivo.distributivo.profesor,
                                                                           periodo).values('id').exists()
                            if not claseactividad and not clase and not actividades:
                                if counter < actividaddetalle.horas:
                                    cls = ClaseActividad(detalledistributivo=detalledistributivo,
                                                         # actividaddetalle=actividaddetalle,
                                                         tipodistributivo=4,
                                                         turno=turno,
                                                         dia=dia,
                                                         inicio=periodo.inicio,
                                                         fin=periodo.fin,
                                                         estadosolicitud=2,
                                                         actividaddetallehorario=actividaddetalle,
                                                         finalizado=True)
                                    cls.save()
                                    counter += 1
                                    response_str += f'<br>{counter}, {detalledistributivo.distributivo.profesor.persona}, {detalledistributivo.distributivo.profesor.persona.cedula}, {actividaddetalle.nombre}'
                                    DEBUG and print(response_str)
                                    if ClaseActividad.objects.filter(actividaddetallehorario=actividaddetalle,
                                                                     detalledistributivo=detalledistributivo,
                                                                     activo=True, status=True).values(
                                            'id').count() == actividaddetalle.horas:
                                        break

            if uncheck:  # Aprueba el horario  solo si ya lo tenía aprobado
                claseactividadestado.estadosolicitud = 2
                claseactividadestado.save()
        notificacion2("Resultados arreglo jefferson", f"{response_str}", persona, None, '/notificacion', persona.pk, 1,
                      'sga', Persona)
    except Exception as ex:
        DEBUG and print(ex.__str__())
        linea_error = 'Error on line {}'.format(sys.exc_info()[-1].tb_lineno)
        notificacion2("Resultados arreglo jefferson", f"{linea_error}. {ex.__str__()}", persona, None, '/notificacion',
                      persona.pk, 1, 'sga', Persona)


def migrar_evidencia_proyecto_vinculacion():
    try:
        from sga.models import Persona, Periodo
        from sga.templatetags.sga_extras import nombremes
        from inno.models import SubactividadDetalleDistributivo, Criterio, MigracionEvidenciaActividad
        from sga.pro_cronograma import CRITERIO_ASOCIADO_VIN, CRITERIO_DIRECTOR_VIN

        ID = 3743
        solicitud = ConfiguracionInformeVinculacion.objects.get(id=ID)
        estadoaprobacion = 2
        ACTIVIDAD_MACRO_VINCULACION = variable_valor('ACTIVIDAD_MACRO_VINCULACION')
        persona = Persona.objects.get(cedula='0928429745')
        periodo = Periodo.objects.get(id=317)
        owner = solicitud.profesor.persona
        lider = solicitud.proyecto.lider()
        es_lider = owner == lider  # Si el dueño de la evidencia es lider del  proyecto
        hoy = datetime.now().date()
        pk_criterio = [151, 150][es_lider]
        pk_subactividad = [CRITERIO_ASOCIADO_VIN, CRITERIO_DIRECTOR_VIN][es_lider]
        es_actividad_macro = SubactividadDetalleDistributivo.objects.values('id').filter(
            actividaddetalledistributivo__criterio__distributivo__periodo=periodo,
            subactividaddocenteperiodo__criterio=pk_subactividad,
            actividaddetalledistributivo__criterio__distributivo__profesor=owner.profesor(), status=True).exists()
        c = CriterioDocencia.objects.get(pk=pk_criterio)
        filtrodetalle = Q(distributivo__profesor__persona=owner, distributivo__periodo=periodo, status=True)
        filtrodetalle &= Q(criteriodocenciaperiodo__criterio=ACTIVIDAD_MACRO_VINCULACION,
                           actividaddetalledistributivo__subactividaddetalledistributivo__subactividaddocenteperiodo__criterio=pk_subactividad)

        if criterio := DetalleDistributivo.objects.filter(filtrodetalle).first():
            subactividaddistributivo = SubactividadDetalleDistributivo.objects.filter(
                actividaddetalledistributivo__criterio=criterio, subactividaddocenteperiodo__criterio__status=True,
                subactividaddocenteperiodo__criterio=pk_subactividad, status=True).first()
            if temp := solicitud.migracionevidenciaactividad_set.filter(status=True).first():
                evidencia = temp.evidencia
            else:
                if not es_actividad_macro:
                    evidencia = solicitud.proyecto.evidenciaactividaddetalledistributivo_set.filter(criterio=criterio,
                                                                                                    desde=solicitud.fecha_inicio,
                                                                                                    hasta=solicitud.fecha_fin,
                                                                                                    status=True,
                                                                                                    generado=True).first()
                else:
                    evidencia = solicitud.proyecto.evidenciaactividaddetalledistributivo_set.filter(
                        subactividad=subactividaddistributivo, criterio=criterio, desde=solicitud.fecha_inicio,
                        hasta=solicitud.fecha_fin, status=True, generado=True).first()

            if not evidencia:
                evidencia = EvidenciaActividadDetalleDistributivo(subactividad=subactividaddistributivo,
                                                                  proyectovinculacion=solicitud.proyecto,
                                                                  criterio=criterio, desde=solicitud.fecha_inicio,
                                                                  hasta=solicitud.fecha_fin,
                                                                  actividad=f'INFORMAR SOBRE LAS ACTIVIDADES REALIZADAS EN EL MES DE {nombremes(solicitud.fecha_fin.month).upper()}',
                                                                  generado=True)
                evidencia.save()
                if not MigracionEvidenciaActividad.objects.filter(evidencia=evidencia, informevinculacion=solicitud,
                                                                  status=True).exists():
                    migracion = MigracionEvidenciaActividad(evidencia=evidencia, informevinculacion=solicitud)
                    migracion.save()

            evidencia.archivo = solicitud.archivo
            evidencia.estadoaprobacion = estadoaprobacion
            historial = HistorialAprobacionEvidenciaActividad(evidencia=evidencia, usuario_creacion=persona.usuario,
                                                              observacion=f"{solicitud.detalle_aprobacion}",
                                                              estadoaprobacion=estadoaprobacion,
                                                              aprobacionpersona=persona if estadoaprobacion == 4 else None,
                                                              fechaaprobacion=hoy if estadoaprobacion == 4 else None)
            historial.save()
            evidencia.save()

    except Exception as ex:
        return False, ex.__str__()


def migrar_evidencias_faltantes_grupos_proyectos():
    try:
        from sga.pro_cronograma import CRITERIO_DIRECTOR_PROYECTO, CRITERIO_CODIRECTOR_PROYECTO_INV, \
            CRITERIO_ASOCIADO_PROYECTO
        from sga.pro_cronograma import CRITERIO_DIRECTOR_GRUPOINVESTIGACION, CRITERIO_INTEGRANTE_GRUPOINVESTIGACION
        from investigacion.models import ProyectoInvestigacionIntegrante
        from inno.models import SubactividadDetalleDistributivo

        EXCLUIR_PROYECTOS_INVESTIGACION = variable_valor('EXCLUIR_PROYECTOS_INVESTIGACION')
        personas = list(ProyectoInvestigacionIntegrante.objects.values_list('persona', flat=True).filter(funcion=1,
                                                                                                         tiporegistro__in=[
                                                                                                             1, 3, 4],
                                                                                                         proyecto__estado_id=37,
                                                                                                         status=True).exclude(
            proyecto__in=EXCLUIR_PROYECTOS_INVESTIGACION))
        personas += list(
            GrupoInvestigacionIntegrante.objects.values_list('persona', flat=True).filter(grupo__vigente=True,
                                                                                          grupo__status=True,
                                                                                          status=True))

        periodo = Periodo.objects.get(id=317)
        for i, detalle in enumerate(DetalleDistributivo.objects.filter(distributivo__profesor__persona__in=personas,
                                                                       distributivo__periodo=periodo,
                                                                       criterioinvestigacionperiodo__criterio__id=68,
                                                                       status=True)):
            SubactividadDetalleDistributivo.objects.filter(actividaddetalledistributivo__criterio=detalle,
                                                           subactividaddocenteperiodo__criterio=CRITERIO_)




    except Exception as ex:
        pass


def llenar_temas_subtemas_semana():
    import sys
    from sga.funciones import notificacion2
    from datetime import datetime, timedelta
    from settings import DEBUG
    from django.db.models import Q
    from sga.funciones import daterange
    from sga.models import Periodo, Persona, TemaAsistencia, SubTemaAsistencia, Leccion, ProfesorMateria, \
        DetalleSilaboSemanalTema, DetalleSilaboSemanalSubtema
    SEMANAS_JUSTIFICADAS = [12, 13, 14, 15]
    id_profesor_materia = [161730]
    periodo = Periodo.objects.get(id=317)
    persona_notifica = Persona.objects.get(id=37121)
    hoy = datetime.now().date()
    try:
        lista_afectados = []
        for profesormateria in ProfesorMateria.objects.filter(id__in=id_profesor_materia).distinct('materia'):
            if silabo := profesormateria.materia.silabo_actual():
                leccion = Leccion.objects.filter(clase__materia=profesormateria.materia, status=True).order_by(
                    'fecha_creacion').first()
                notificacion2("Resultados arreglo jefferson",
                              f"Silabo semanal {silabo.silabosemanal_set.filter(status=True, fechainiciosemana__lte=hoy, numsemana__in=SEMANAS_JUSTIFICADAS)}",
                              persona_notifica, None, '/notificacion', persona_notifica.pk, 1, 'sga', Persona)
                for ss in silabo.silabosemanal_set.filter(status=True, fechainiciosemana__lte=hoy,
                                                          numsemana__in=SEMANAS_JUSTIFICADAS):
                    if __leccion := Leccion.objects.filter(clase__materia=profesormateria.materia,
                                                           fecha__gte=ss.fechainiciosemana,
                                                           fecha__lte=ss.fechafinciosemana, status=True).order_by(
                            'fecha_creacion').first():
                        leccion = __leccion
                    filtro_silabos_1_y_2 = Q(silabosemanal__silabo=silabo, silabosemanal=ss, status=True)

                    # No se excluyen los temas marca2 porque pueden haber temas sin subtemas marca2
                    for tema in DetalleSilaboSemanalTema.objects.filter(filtro_silabos_1_y_2):
                        fecharegistro = leccion.fecha
                        fi, ff = tema.silabosemanal.fechainiciosemana, tema.silabosemanal.fechafinciosemana + timedelta(
                            1)
                        if not fi <= fecharegistro <= ff:
                            for __date in daterange(fi, ff):
                                if __date.isoweekday() == leccion.clase.dia:
                                    fecharegistro = __date

                        if temaasistencia := TemaAsistencia.objects.filter(
                                leccion__clase__materia=profesormateria.materia, tema=tema, status=True):
                            temaasistencia.update(fecha=fecharegistro)
                            temaasistencia = temaasistencia.first()
                        else:
                            temaasistencia = TemaAsistencia(leccion=leccion, tema=tema, fecha=fecharegistro)
                            temaasistencia.save()

                        registro_subtema = 0
                        # Se excluyen los subtemas que ya fueron registrados en la materia
                        __exclude = SubTemaAsistencia.objects.filter(fecha__lte=ff + timedelta(weeks=2),
                                                                     tema__leccion__clase__materia=profesormateria.materia,
                                                                     status=True).values_list('subtema', flat=True)
                        for subtema in DetalleSilaboSemanalSubtema.objects.filter(
                                filtro_silabos_1_y_2):  # .exclude(id__in=__exclude):
                            if subtemaasistencia := SubTemaAsistencia.objects.filter(tema=temaasistencia,
                                                                                     subtema=subtema, status=True):
                                subtemaasistencia.update(fecha=fecharegistro)
                            else:
                                subtemaasistencia = SubTemaAsistencia(tema=temaasistencia, subtema=subtema,
                                                                      fecha=fecharegistro)
                                subtemaasistencia.save()

                            registro_subtema += 1
                        registro_subtema and lista_afectados.append(
                            f"{registro_subtema} - SEMANA {ss.numsemana} - {profesormateria} - fecha registro {fecharegistro}")
                    DEBUG and print(f"{profesormateria}")
        lista_afectados and notificacion2("Resultados arreglo jefferson", f" <br>".join(lista_afectados),
                                          persona_notifica, None, '/notificacion', persona_notifica.pk, 1, 'sga',
                                          Persona)
    except Exception as ex:
        DEBUG and print(ex.__str__())
        linea_error = 'Error on line {}'.format(sys.exc_info()[-1].tb_lineno)
        notificacion2("Resultados arreglo jefferson", f"{linea_error}. {ex.__str__()}", persona_notifica, None,
                      '/notificacion', persona_notifica.pk, 1, 'sga', Persona)


def editar_fecha_registro_temas_subtemas():
    from sga.funciones import notificacion2
    from sga.models import TemaAsistencia, SubTemaAsistencia, LeccionGrupo
    persona_notifica = Persona.objects.get(id=37121)  # Jefferson Cuadrado
    try:
        for tema in TemaAsistencia.objects.filter(leccion__clase__materia__nivel__periodo=317, usuario_creacion_id=1,
                                                  tema__silabosemanal__numsemana__in=(1, 2), status=True):
            if not tema.leccion.fecha == tema.fecha:
                tema.fecha = tema.leccion.fecha
                tema.save()

        for subtema in SubTemaAsistencia.objects.filter(tema__leccion__clase__materia__nivel__periodo=317,
                                                        usuario_creacion_id=1,
                                                        tema__tema__silabosemanal__numsemana__in=(1, 2), status=True):
            if not subtema.tema.leccion.fecha == subtema.fecha:
                subtema.fecha = subtema.tema.leccion.fecha
                subtema.save()

        lecciongrupo = LeccionGrupo.objects.db_manager("sga_select").get(pk=1530527)
        for tema in TemaAsistencia.objects.filter(tema=810427, status=True):
            if not tema.fecha == lecciongrupo.fecha:
                tema.fecha = lecciongrupo.fecha
                tema.save()

        for subtema in SubTemaAsistencia.objects.filter(subtema=1943122, status=True):
            if not lecciongrupo.fecha == subtema.fecha:
                subtema.fecha = lecciongrupo.fecha
                subtema.save()

        notificacion2("Resultados arreglo editar_fecha_registro_temas_subtemas()",
                      f"Proceso terminado satisfactoriamente...", persona_notifica, None, '/notificacion',
                      persona_notifica.pk, 1, 'sga', Persona)
    except Exception as ex:
        notificacion2("Resultados arreglo jefferson", f"{ex.__str__()}", persona_notifica, None, '/notificacion',
                      persona_notifica.pk, 1, 'sga', Persona)


def actualizar_horas_planificadas_bitacora():
    from sga.funciones import notificacion2
    from sga.models import Persona
    from django.db.models import Sum, F, ExpressionWrapper, TimeField
    persona_notifica = Persona.objects.get(id=37121)
    try:
        personas, lista_afectados = [45821, 48257, 44111, 147952, 255242, 36266, 21769, 27715], []
        for bitacora in BitacoraActividadDocente.objects.filter(fechafin__month=4, fechafin__year=2024,
                                                                criterio__distributivo__profesor__persona__in=personas,
                                                                status=True).distinct():
            listadodetalle = bitacora.detallebitacoradocente_set.filter(status=True).annotate(
                diferencia=ExpressionWrapper(F('horafin') - F('horainicio'), output_field=TimeField())).order_by(
                'fecha', 'horainicio', 'horafin')
            totalhorasaprobadas, planificadas_old = 0, bitacora.horasplanificadas
            if th := listadodetalle.filter(estadoaprobacion=2, status=True).aggregate(total=Sum('diferencia'))['total']:
                horas, minutos = (th.total_seconds() / 3600).__str__().split('.')
                totalhorasaprobadas = float("%s.%s" % (horas, round(float('0.' + minutos) * 60)))
                if totalhorasaprobadas == bitacora.get_horasregistradas():
                    bitacora.horasplanificadas = totalhorasaprobadas
                    bitacora.save()
                    DEBUG and print(f"({bitacora.pk}, {bitacora.criterio.distributivo.profesor.persona}), ")
                    lista_afectados.append(
                        f'({bitacora.pk}, {planificadas_old}, {bitacora.horasplanificadas}, {bitacora.criterio.distributivo.profesor.persona}),<br>')
        lista_afectados and notificacion2("Resultados arreglo jefferson", f" <br>".join(lista_afectados),
                                          persona_notifica, None, '/notificacion', persona_notifica.pk, 1, 'sga',
                                          Persona)
    except Exception as ex:
        DEBUG and print(ex.__str__())
        linea_error = 'Error on line {}'.format(sys.exc_info()[-1].tb_lineno)
        notificacion2("Resultados arreglo jefferson", f"{linea_error}. {ex.__str__()}", persona_notifica, None,
                      '/notificacion', persona_notifica.pk, 1, 'sga', Persona)


def arreglo_jefferson():
    from sagest.models import AnioEjercicio
    from sga.funciones import notificacion2, null_to_decimal
    from sga.pro_cronograma import CRITERIO_ELABORA_ARTICULO_CIENTIFICO
    from inno.models import SecuenciaInformeMensualActividades, MigracionEvidenciaActividad
    from sga.models import Persona, Profesor, ConfiguracionInformePracticasPreprofesionales, InformeMensualDocentesPPP, \
        EvidenciaActividadDetalleDistributivo
    persona_notifica = Persona.objects.get(id=37121)
    try:
        hoy = datetime.now().date()
        profesor = Profesor.objects.get(persona=320799)
        configuracion = ConfiguracionInformePracticasPreprofesionales.objects.get(id=1367)

        informeppp = InformeMensualDocentesPPP.objects.filter(persona=profesor, mes=configuracion.mes,
                                                              anio=configuracion.anio, status=True).first()
        aModel = AnioEjercicio.objects.filter(anioejercicio=configuracion.anio, status=True).first()
        anioejercicio = AnioEjercicio.objects.create(anioejercicio=configuracion.anio) if not aModel else aModel
        secuencia = null_to_decimal(
            InformeMensualDocentesPPP.objects.filter(persona=profesor, status=True, secuencia__status=True,
                                                     secuencia__anioejercicio=anioejercicio,
                                                     secuencia__tipo=1).aggregate(valor=Max('secuencia__secuencia'))[
                'valor'])
        numeracion = '%03d' % ((1 + secuencia) if not informeppp else informeppp.secuencia.secuencia)

        secuencia = SecuenciaInformeMensualActividades(tipo=1, secuencia=numeracion, anioejercicio=anioejercicio)
        secuencia.save()

        informe = InformeMensualDocentesPPP(persona=profesor, mes=configuracion.mes, anio=configuracion.anio,
                                            fechageneracion=hoy, secuencia=secuencia, configuracion=configuracion)
        informe.save()

        evidencia = EvidenciaActividadDetalleDistributivo.objects.get(id=107933)
        evidencia.informe = informe
        evidencia.save()

        # Dirigir proyecto de investigacion
        evidencia_id = 106328 if DEBUG else 107590
        evidencia = EvidenciaActividadDetalleDistributivo.objects.filter(id=evidencia_id).values()[0]
        evidencia.pop('id')
        evidencia['criterio_id'] = 187994
        evidencia['subactividad_id'] = 7358
        ev = EvidenciaActividadDetalleDistributivo(**evidencia)
        ev.save()

        m = MigracionEvidenciaActividad.objects.filter(evidencia=evidencia_id, status=True).first()
        MigracionEvidenciaActividad(evidencia=ev, evidenciabase=m.evidencia,
                                    proyectoinvestigacion=m.proyectoinvestigacion).save()

        for anexo in AnexoEvidenciaActividad.objects.filter(evidencia=evidencia_id):
            AnexoEvidenciaActividad(evidencia=ev, observacion=anexo.observacion, archivo=anexo.archivo).save()

        # Dirigir grupo de investigacion
        evidencias = [106122, 105895] if DEBUG else [107916, 107379]
        for evd in evidencias:
            evidencia = EvidenciaActividadDetalleDistributivo.objects.filter(id=evd).values()[0]
            evidencia.pop('id')
            evidencia['criterio_id'] = 187994
            evidencia['subactividad_id'] = 7357
            ev = EvidenciaActividadDetalleDistributivo(**evidencia)
            ev.save()

            MigracionEvidenciaActividad(evidencia=ev, evidenciabase_id=evd).save()

            for anexo in AnexoEvidenciaActividad.objects.filter(evidencia=evidencia_id):
                AnexoEvidenciaActividad(evidencia=ev, observacion=anexo.observacion, archivo=anexo.archivo).save()

        # Aprobar las bitacoras de escribir aticulo cientifico
        for bitacora in BitacoraActividadDocente.objects.filter(detallebitacoradocente__status=True,
                                                                subactividad__subactividaddocenteperiodo__criterio__id=CRITERIO_ELABORA_ARTICULO_CIENTIFICO,
                                                                status=True).distinct():
            bitacora.detallebitacoradocente_set.filter(status=True).update(estadoaprobacion=2)
            bitacora.estadorevision = 3
            bitacora.save()

        notificacion2("Resultados arreglo jefferson", f"Proceso finalizado con exito...", persona_notifica, None,
                      '/notificacion', persona_notifica.pk, 1, 'sga', Persona)
    except Exception as ex:
        DEBUG and print(ex.__str__())
        linea_error = 'Error on line {}'.format(sys.exc_info()[-1].tb_lineno)
        notificacion2("Resultados arreglo jefferson", f"{linea_error}. {ex.__str__()}", persona_notifica, None,
                      '/notificacion', persona_notifica.pk, 1, 'sga', Persona)


def migrar_evidencia_investigacion():
    try:
        from inno.models import MigracionEvidenciaActividad
        from investigacion.models import BitacoraActividadDocente
        from sga.pro_cronograma import CRITERIO_ELABORA_ARTICULO_CIENTIFICO
        for bitacora in BitacoraActividadDocente.objects.filter(detallebitacoradocente__status=True,
                                                                subactividad__subactividaddocenteperiodo__criterio__id=CRITERIO_ELABORA_ARTICULO_CIENTIFICO,
                                                                status=True).distinct():
            bitacora.detallebitacoradocente_set.filter(status=True).update(estadoaprobacion=2)
            bitacora.estadorevision = 3
            bitacora.save()
    except Exception as ex:
        pass


def migrar_evidencia_investigacion_informe():
    from investigacion.models import ProyectoInvestigacionIntegrante
    from inno.models import SubactividadDetalleDistributivo, MigracionEvidenciaActividad
    from sga.models import Persona, EvidenciaActividadDetalleDistributivo, HistorialAprobacionEvidenciaActividad, \
        AnexoEvidenciaActividad, EvidenciaActividadAudi
    from sga.pro_cronograma import CRITERIO_CODIRECTOR_PROYECTO_INV, CRITERIO_DIRECTOR_PROYECTO, \
        CRITERIO_ASOCIADO_PROYECTO, CRITERIO_INTEGRANTE_GRUPOINVESTIGACION
    from sga.funciones import notificacion2
    persona_notifica = Persona.objects.get(id=37121)
    try:
        CRITERIO_ACTIVIDAD_MACRO = 68
        for evidencia in EvidenciaActividadDetalleDistributivo.objects.filter(id__in=(106763, 108047)):
            periodo = evidencia.criterio.distributivo.periodo
            persona = evidencia.criterio.distributivo.profesor.persona
            funcion = 0
            if evidencia.subactividad.subactividaddocenteperiodo.criterio.pk == CRITERIO_CODIRECTOR_PROYECTO_INV:
                funcion = 2
            if evidencia.subactividad.subactividaddocenteperiodo.criterio.pk == CRITERIO_DIRECTOR_PROYECTO:
                funcion = 1
            if inv := evidencia.criterio.criterioinvestigacionperiodo:
                can_upload_evidence, exclude_ = False, [1]
                if lider := ProyectoInvestigacionIntegrante.objects.filter(tiporegistro__in=[1, 3, 4], persona=persona,
                                                                           proyecto__estado_id=37, funcion=funcion,
                                                                           status=True).exclude(
                        proyecto__cerrado=True).first():
                    proyecto = lider.proyecto
                    if inv.es_actividadmacro:
                        # Cambios para actividad macro
                        lider_proyecto = lider.proyecto.integrantes_proyecto().filter(funcion=1).first()
                        if not SubactividadDetalleDistributivo.objects.filter(
                                subactividaddocenteperiodo__criterio=CRITERIO_DIRECTOR_PROYECTO,
                                actividaddetalledistributivo__criterio__distributivo__profesor__persona=lider_proyecto.persona,
                                actividaddetalledistributivo__criterio__distributivo__periodo=periodo,
                                status=True).exists():
                            can_upload_evidence, exclude_ = True, [1, 2]

                        if (
                                evidencia.subactividad and evidencia.subactividad.subactividaddocenteperiodo.criterio.pk == CRITERIO_DIRECTOR_PROYECTO) or can_upload_evidence:
                            if not MigracionEvidenciaActividad.objects.values('id').filter(evidencia=evidencia,
                                                                                           proyectoinvestigacion=lider.proyecto,
                                                                                           status=True).exists():
                                m = MigracionEvidenciaActividad(evidencia=evidencia,
                                                                proyectoinvestigacion=lider.proyecto)
                                m.save()
                            for integrante in lider.proyecto.integrantes_proyecto().exclude(funcion__in=exclude_):
                                _evidencia = None
                                integrante_profesor = integrante.profesor if integrante.profesor else Profesor.objects.filter(
                                    persona=integrante.persona).first()
                                if distributivo := DetalleDistributivo.objects.filter(
                                        criterioinvestigacionperiodo__criterio__id=CRITERIO_ACTIVIDAD_MACRO,
                                        distributivo__profesor=integrante_profesor,
                                        distributivo__periodo=evidencia.criterio.distributivo.periodo,
                                        status=True).first():
                                    CRITERIO = CRITERIO_ASOCIADO_PROYECTO
                                    if integrante.funcion == 2:
                                        CRITERIO = CRITERIO_CODIRECTOR_PROYECTO_INV
                                    if subactividad := SubactividadDetalleDistributivo.objects.filter(
                                            actividaddetalledistributivo__vigente=True,
                                            actividaddetalledistributivo__criterio=distributivo,
                                            subactividaddocenteperiodo__criterio=CRITERIO,
                                            subactividaddocenteperiodo__actividad__status=True,
                                            subactividaddocenteperiodo__actividad__criterio__status=True,
                                            subactividaddocenteperiodo__status=True,
                                            subactividaddocenteperiodo__criterio__status=True,
                                            actividaddetalledistributivo__status=True, status=True).first():
                                        if migracion := MigracionEvidenciaActividad.objects.filter(
                                                evidencia__subactividad=subactividad, evidenciabase=evidencia.pk,
                                                status=True).first():
                                            _evidencia = migracion.evidencia
                                            _evidencia.actividaddetalledistributivo = evidencia.actividaddetalledistributivo
                                            _evidencia.desde = evidencia.desde
                                            _evidencia.hasta = evidencia.hasta
                                            _evidencia.actividad = evidencia.actividad
                                            _evidencia.aprobado = evidencia.aprobado
                                            _evidencia.archivo = evidencia.archivo
                                            _evidencia.usuarioaprobado = evidencia.usuarioaprobado
                                            _evidencia.fechaaprobado = evidencia.fechaaprobado
                                            _evidencia.estadoaprobacion = evidencia.estadoaprobacion
                                            _evidencia.archivofirmado = evidencia.archivofirmado
                                            _evidencia.save()
                                        else:
                                            _dict = EvidenciaActividadDetalleDistributivo.objects.filter(
                                                pk=evidencia.pk).values()[0]
                                            _dict.pop('id')
                                            _evidencia = EvidenciaActividadDetalleDistributivo(**_dict)
                                            _evidencia.criterio = distributivo
                                            _evidencia.subactividad = subactividad
                                            _evidencia.save()
                                            m = MigracionEvidenciaActividad(evidencia=_evidencia,
                                                                            evidenciabase=evidencia,
                                                                            proyectoinvestigacion=lider.proyecto)
                                            m.save()
                                        _evidencia.anexoevidenciaactividad_set.filter(status=True).delete()
                                        for anexo in evidencia.anexoevidenciaactividad_set.filter(status=True):
                                            a = AnexoEvidenciaActividad(evidencia=_evidencia,
                                                                        observacion=anexo.observacion,
                                                                        archivo=anexo.archivo)
                                            a.save()
                                        _evidencia.evidenciaactividadaudi_set.filter(status=True).delete()
                                        for anexo in evidencia.evidenciaactividadaudi_set.filter(status=True):
                                            a = EvidenciaActividadAudi(evidencia=_evidencia, archivo=anexo.archivo)
                                            a.save()
                                        _evidencia.historialaprobacionevidenciaactividad_set.filter(
                                            status=True).delete()
                                        for anexo in evidencia.historialaprobacionevidenciaactividad_set.filter(
                                                status=True):
                                            a = HistorialAprobacionEvidenciaActividad(evidencia=_evidencia,
                                                                                      aprobacionpersona=anexo.aprobacionpersona,
                                                                                      observacion=anexo.observacion,
                                                                                      fechaaprobacion=anexo.fechaaprobacion,
                                                                                      estadoaprobacion=anexo.estadoaprobacion)
                                            a.save()

        notificacion2("Resultados arreglo jefferson", 'Arreglo jefferson migrar investigacion ejecutado...',
                      persona_notifica, None, '/notificacion', persona_notifica.pk, 1, 'sga', Persona)
    except Exception as ex:
        DEBUG and print(ex.__str__())
        linea_error = 'Error on line {}'.format(sys.exc_info()[-1].tb_lineno)
        notificacion2("Resultados arreglo jefferson", f"{linea_error}. {ex.__str__()}", persona_notifica, None,
                      '/notificacion', persona_notifica.pk, 1, 'sga', Persona)


def arreglo_jefferson_v2():
    from sagest.models import AnioEjercicio
    from sga.funciones import notificacion2, null_to_decimal
    from sga.pro_cronograma import CRITERIO_ELABORA_ARTICULO_CIENTIFICO
    from inno.models import SecuenciaInformeMensualActividades, MigracionEvidenciaActividad
    from sga.models import Persona, Profesor, ConfiguracionInformePracticasPreprofesionales, InformeMensualDocentesPPP, \
        EvidenciaActividadDetalleDistributivo
    persona_notifica = Persona.objects.get(id=37121)
    try:
        hoy = datetime.now().date()
        profesor = Profesor.objects.get(persona=320799)
        configuracion = ConfiguracionInformePracticasPreprofesionales.objects.get(id=1367)

        informeppp = InformeMensualDocentesPPP.objects.filter(persona=profesor, mes=configuracion.mes,
                                                              anio=configuracion.anio, status=True).first()
        aModel = AnioEjercicio.objects.filter(anioejercicio=configuracion.anio, status=True).first()
        anioejercicio = AnioEjercicio.objects.create(anioejercicio=configuracion.anio) if not aModel else aModel
        secuencia = null_to_decimal(
            InformeMensualDocentesPPP.objects.filter(persona=profesor, status=True, secuencia__status=True,
                                                     secuencia__anioejercicio=anioejercicio,
                                                     secuencia__tipo=1).aggregate(valor=Max('secuencia__secuencia'))[
                'valor'])
        numeracion = '%03d' % ((1 + secuencia) if not informeppp else informeppp.secuencia.secuencia)

        secuencia = SecuenciaInformeMensualActividades(tipo=1, secuencia=numeracion, anioejercicio=anioejercicio)
        secuencia.save()

        informe = InformeMensualDocentesPPP(persona=profesor, mes=configuracion.mes, anio=configuracion.anio,
                                            fechageneracion=hoy, secuencia=secuencia, configuracion=configuracion)
        informe.save()

        evidencia = EvidenciaActividadDetalleDistributivo.objects.get(id=107933)
        evidencia.informe = informe
        evidencia.save()

        # Aprobar las bitacoras de escribir aticulo cientifico
        for bitacora in BitacoraActividadDocente.objects.filter(detallebitacoradocente__status=True,
                                                                subactividad__subactividaddocenteperiodo__criterio__id=CRITERIO_ELABORA_ARTICULO_CIENTIFICO,
                                                                status=True).distinct():
            bitacora.detallebitacoradocente_set.filter(status=True).update(estadoaprobacion=2)
            bitacora.estadorevision = 3
            bitacora.save()

        notificacion2("Resultados arreglo jefferson", f"Proceso finalizado con exito...", persona_notifica, None,
                      '/notificacion', persona_notifica.pk, 1, 'sga', Persona)
    except Exception as ex:
        DEBUG and print(ex.__str__())
        linea_error = 'Error on line {}'.format(sys.exc_info()[-1].tb_lineno)
        notificacion2("Resultados arreglo arreglo_jefferson_v2", f"{linea_error}. {ex.__str__()}", persona_notifica,
                      None, '/notificacion', persona_notifica.pk, 1, 'sga', Persona)


def vaciar_tablas():
    if DEBUG:
        cursor = connections['sga_select'].cursor()
        lista_tablas_criticas = ["sga_sesionzoom", "sga_respuestaevaluacionacreditacion",
                                 "sga_matriculaseguimientotutor", "sga_evaluaciongenerica", "sga_datorespuestaencuesta",
                                 "sga_correomatriculaseguimientotutor", "sga_bibliograbiaapasilabo",
                                 "sga_auditorianotas", "sga_asistenciamoodle", "sga_actividadessakaialumno",
                                 "sagest_registromarcada", "sagest_logrubros", "sagest_logmarcada", "sagest_logdia",
                                 "sagest_factura", "inno_solicitudtutoriaindividual",
                                 "inno_materiaasignadaplanificacionsedevirtualexamen",
                                 "inno_calendariorecursoactividadalumno",
                                 "inno_calendariorecursoactividadalumnomotificacion",
                                 "sga_correomatriculatutornotificacion", "bd_logentrylogin", "django_admin_log",
                                 "sga_notificacion", "django_admin_log_backup", "django_admin_log_backupdos",
                                 "bd_geolocation", "bd_geolocationuser"]
        for table in lista_tablas_criticas:
            sql = f"TRUNCATE TABLE {table} CASCADE"
            cursor.execute(sql)

            print(f'{table} is empty...')


def del_bitacora_duplicada():
    try:
        from investigacion.models import BitacoraActividadDocente
        BitacoraActividadDocente.objects.get(id=2670).delete()
    except Exception as ex:
        pass


def reporte_marcadas(**kwargs):
    import io
    import xlsxwriter
    from django.http import HttpResponse
    periodo = Periodo.objects.get(id=kwargs.pop('periodo'))
    profesor = Profesor.objects.get(id=kwargs.pop('profesor'))
    try:
        __author__ = 'Unemi'
        ahora = datetime.now()
        time_codigo = ahora.strftime('%Y%m%d%H%M%S')
        name_file = f'{time_codigo}_reporte_marcadas_{profesor.pk}.xlsx'

        # output = io.BytesIO()
        # workbook = xlsxwriter.Workbook(output)

        output_folder = f"{BASE_DIR}/media/{name_file}"
        workbook = xlsxwriter.Workbook(output_folder)

        fuentecabecera = workbook.add_format({'align': 'center', 'bg_color': 'silver', 'border': 1, 'bold': 1})
        formatoceldacenter = workbook.add_format({'border': 1, 'valign': 'vcenter', 'align': 'center'})
        formatoceldacenter_bold = workbook.add_format({'border': 1, 'valign': 'vcenter', 'align': 'center', 'bold': 1})
        formatoceldaleft = workbook.add_format({'border': 1, 'valign': 'vleft', 'align': 'left'})
        formatoceldaleft_bold = workbook.add_format({'border': 1, 'valign': 'vcenter', 'align': 'left', 'bold': 1})
        fuenteencabezado = workbook.add_format(
            {'align': 'center', 'bg_color': '#1C3247', 'font_color': 'white', 'border': 1, 'font_size': 24, 'bold': 1})

        for profesormateria in ProfesorMateria.objects.filter(profesor=profesor, materia__nivel__periodo=periodo,
                                                              materia__status=True).distinct('materia').order_by(
                'materia', 'materia__asignaturamalla__asignatura__nombre'):
            alias = profesormateria.materia.alias if profesormateria.materia.alias else f"{profesormateria.materia.asignaturamalla.asignatura.nombre}"[
                                                                                        :-10]
            nombrepagina = f"{alias}_{profesormateria.materia.paralelo}".replace(' ', '_').lower()
            ws = workbook.add_worksheet(nombrepagina)
            columnas = [(u"Semana", 20), (u"Tema", 80), (u"Subtema", 80), (u"¿Marcada?", 10), (u"Fecha Marcada", 50)]
            ws.merge_range(0, 0, 0, len(columnas) - 1, 'UNIVERSIDAD ESTATAL DE MILAGRO', fuenteencabezado)
            ws.merge_range(1, 0, 1, len(columnas) - 1, f'{profesormateria.materia}', fuenteencabezado)
            row_num, numcolum = 3, 0
            for col_name in columnas:
                ws.write(row_num, numcolum, col_name[0], fuentecabecera)
                ws.set_column(numcolum, numcolum, col_name[1])
                numcolum += 1
            row_num += 1
            ws.merge_range(row_num, 1, row_num, len(columnas) - 1, f'{profesormateria.materia}',
                           formatoceldacenter_bold)
            row_num += 1
            if silabo := Silabo.objects.filter(profesor=profesormateria.profesor, materia=profesormateria.materia,
                                               status=True).first():
                silabossemanales = silabo.silabosemanal_set.filter(fechainiciosemana__lte=ahora, status=True)
                for silabosemanal in silabossemanales:
                    temas = silabosemanal.detallesilabosemanaltema_set.filter(
                        temaunidadresultadoprogramaanalitico__status=True, status=True).distinct()
                    num_subtemas = silabosemanal.detallesilabosemanalsubtema_set.filter(
                        Q(subtemaunidadresultadoprogramaanalitico__temaunidadresultadoprogramaanalitico__in=temas.values_list(
                            'temaunidadresultadoprogramaanalitico', flat=True),
                          subtemaunidadresultadoprogramaanalitico__status=True,
                          subtemaunidadresultadoprogramaanalitico__temaunidadresultadoprogramaanalitico__isnull=False,
                          subtemaunidadresultadoprogramaanalitico__temaunidadresultadoprogramaanalitico__status=True,
                          status=True)).distinct().count()
                    last_row = (row_num + num_subtemas) - 1
                    if (row_num - last_row) in [0, 1]:
                        ws.write(row_num, 0,
                                 f'Semana {silabosemanal.numsemana}: {silabosemanal.fechainiciosemana} - {silabosemanal.fechafinciosemana}',
                                 formatoceldacenter_bold)
                    else:
                        ws.merge_range(row_num, 0, last_row, 0, f'Semana {silabosemanal.numsemana}', formatoceldacenter)
                    for tema in temas:
                        subtemasportema = silabosemanal.detallesilabosemanalsubtema_set.filter(
                            Q(subtemaunidadresultadoprogramaanalitico__temaunidadresultadoprogramaanalitico=tema.temaunidadresultadoprogramaanalitico,
                              subtemaunidadresultadoprogramaanalitico__status=True,
                              subtemaunidadresultadoprogramaanalitico__temaunidadresultadoprogramaanalitico__isnull=False,
                              subtemaunidadresultadoprogramaanalitico__temaunidadresultadoprogramaanalitico__status=True,
                              status=True)).distinct()
                        last_row = row_num + len(subtemasportema) - 1
                        if (row_num - last_row) in [0, 1]:
                            ws.write(row_num, 1, f'{tema.temaunidadresultadoprogramaanalitico}', formatoceldaleft_bold)
                        else:
                            ws.merge_range(row_num, 1, last_row, 1, f'{tema.temaunidadresultadoprogramaanalitico}',
                                           formatoceldaleft_bold)
                        for subtema in subtemasportema:
                            ws.write(row_num, 2, f"{subtema.subtemaunidadresultadoprogramaanalitico}", formatoceldaleft)
                            if marcada := subtema.subtemaasistencia_set.filter(status=True).first():
                                ws.write(row_num, 3, f"Si", formatoceldacenter)
                                ws.write(row_num, 4, f"{marcada.fecha}", formatoceldacenter)
                            else:
                                ws.write(row_num, 3, f"No", formatoceldacenter)
                                ws.write(row_num, 4, f"-", formatoceldacenter)
                            row_num += 1
                    if not temas:
                        ws.merge_range(row_num, 1, row_num, len(columnas) - 1, f'{silabosemanal}',
                                       formatoceldacenter_bold)
                        row_num += 1

        workbook.close()
        # output.seek(0)
        # response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        # response['Content-Disposition'] = 'attachment; filename=mi_archivo_excel.xlsx'
        # return response
    except Exception as ex:
        pass


def actualizar_estado_evidencia_vinculacion():
    from sga.models import ConfiguracionInformeVinculacion
    try:
        for con in ConfiguracionInformeVinculacion.objects.filter(id__in=[3958, 3537, 3581]):
            if migracion := con.migracionevidenciaactividad_set.filter(status=True).first():
                evidencia = migracion.evidencia
                con.archivo = evidencia.archivofirmado if evidencia.archivofirmado else evidencia.archivo
                con.status = True
                con.save()
    except Exception as ex:
        pass


def reporte_carreras_nivelacion_poa():
    try:
        import xlsxwriter

        __author__ = 'Unemi'
        ahora = datetime.now()
        time_codigo = ahora.strftime('%Y%m%d%H%M%S')
        name_file = f'000_eficiencia_curso_nivelacion_{time_codigo}.xlsx'

        output_folder = f"{BASE_DIR}/media/arreglo_jefferson/{name_file}"
        workbook = xlsxwriter.Workbook(output_folder)
        ws = workbook.add_worksheet('resultados')

        anioconsulta = 2023
        filtroperiodo = Q(inicio__year=anioconsulta, visible=True, status=True)
        if not anioconsulta >= 2023:
            filtroperiodo &= Q(inicio__year=anioconsulta, clasificacion=3)
        else:
            filtroperiodo &= Q(tipo=2, matriculacionactiva=True, clasificacion=1)

        periodos = Periodo.objects.filter(filtroperiodo)
        carreras = Materia.objects.values_list('asignaturamalla__malla__carrera', flat=True).filter(
            nivel__periodo__in=periodos, asignaturamalla__malla__carrera__coordinacion=9, status=True,
            nivel__status=True, asignaturamalla__status=True, asignaturamalla__malla__status=True).distinct()

        fuentecabecera = workbook.add_format({'align': 'center', 'bg_color': 'silver', 'border': 1, 'bold': 1})
        formatoceldacenter = workbook.add_format({'border': 1, 'valign': 'vcenter', 'align': 'center'})
        formatoceldacenter_bold = workbook.add_format({'border': 1, 'valign': 'vcenter', 'align': 'center', 'bold': 1})
        formatoceldaleft = workbook.add_format({'border': 1, 'valign': 'vleft', 'align': 'left'})
        formatoceldaleft_bold = workbook.add_format({'border': 1, 'valign': 'vcenter', 'align': 'left', 'bold': 1})
        fuenteencabezado = workbook.add_format(
            {'align': 'center', 'bg_color': '#1C3247', 'font_color': 'white', 'border': 1, 'font_size': 24, 'bold': 1})

        columnas = [(u"#", 10),
                    (u"Carrera", 90),
                    (u"Aprobaron Nivelacion 1ra Matricula", 30),
                    (f"Aprobaron Nivelacion 2da Matrícula {anioconsulta}", 40),
                    (f"Iniciaron Nivelación en el Período {anioconsulta}", 40),
                    (f"Porcentaje Total", 20)
                    ]
        ws.merge_range(0, 0, 0, len(columnas) - 1, 'UNIVERSIDAD ESTATAL DE MILAGRO', fuenteencabezado)
        ws.merge_range(1, 0, 1, len(columnas) - 1, f'Reporte de eficiencia del curso de nivelación de la carrera',
                       fuenteencabezado)
        row_num, numcolum = 3, 0
        for col_name in columnas:
            ws.write(row_num, numcolum, col_name[0], fuentecabecera)
            ws.set_column(numcolum, numcolum, col_name[1])
            numcolum += 1
        row_num += 1

        nombreperiodos = ''
        for i, periodo in enumerate(periodos):
            nombreperiodos += f"{periodo.inicio.strftime('%d/%m/%Y')} - {periodo.fin.strftime('%d/%m/%Y')}"
            if not i + 1 == len(periodos):
                nombreperiodos += ' | '

        ws.merge_range(2, 0, 2, len(columnas) - 1, f'{nombreperiodos}', fuenteencabezado)
        for counter, carrera in enumerate(Carrera.objects.filter(id__in=carreras, activa=True, status=True)):
            filtrobase = Q(materia__asignaturamalla__malla__carrera=carrera, retiramateria=False, materia__status=True,
                           matricula__inscripcion__persona__real=True, status=True)
            ws.write(row_num, 0, f"{counter + 1}", formatoceldacenter)
            ws.write(row_num, 1, f"{carrera}", formatoceldaleft)

            # Número de estudiantes que aprobaron nivelación de la carrera en primera matricula en el periodo 2024
            aprobaronnivelacionprimeramatricula = MateriaAsignada.objects.values('id').filter(
                filtrobase & Q(materia__cerrado=True, matriculas=1, materia__nivel__periodo__in=periodos,
                               matricula__aprobado=True)).order_by('matricula__inscripcion__persona__cedula').distinct(
                'matricula__inscripcion__persona__cedula').count()

            # Número de estudiantes que aprobaron nivelación de la carrera en segunda matricula en el período 2024
            aprobaronnivelacionsegundamatricula2024 = MateriaAsignada.objects.values('id').filter(
                filtrobase & Q(materia__cerrado=True, matriculas=2, materia__nivel__periodo__in=periodos,
                               matricula__aprobado=True)).order_by('matricula__inscripcion__persona__cedula').distinct(
                'matricula__inscripcion__persona__cedula').count()

            # Número de estudiantes de la carrera que iniciaron nivelación en el período 2024
            estudiantesiniciaronnivelacionperiodo2024 = MateriaAsignada.objects.values('id').filter(
                filtrobase & Q(materia__nivel__periodo__in=periodos)).order_by(
                'matricula__inscripcion__persona__cedula').distinct('matricula__inscripcion__persona__cedula').count()

            # (estudiantes que aprobaron nivelación de la carrera en primera matricula + 0,5 * estudiantes que aprobaron nivelación de la carrera en segunda matricula en el período 2024 / estudiantes de la carrera que iniciaron nivelación en el período 2024) * 100
            a, b, c = aprobaronnivelacionprimeramatricula, aprobaronnivelacionsegundamatricula2024, estudiantesiniciaronnivelacionperiodo2024
            porcentajetotal = (((a + (0.5 * b)) / c) * 100)  # (a + 0.5 * b / c) * 100

            ws.write(row_num, 2, f"{aprobaronnivelacionprimeramatricula}", formatoceldacenter)
            ws.write(row_num, 3, f"{aprobaronnivelacionsegundamatricula2024}", formatoceldacenter)
            ws.write(row_num, 4, f"{estudiantesiniciaronnivelacionperiodo2024}", formatoceldacenter)
            ws.write(row_num, 5, f"{porcentajetotal:.2f} %", formatoceldacenter)

            # for x in MatriculacionPrimerNivelCarrera.objects.filter(carreraadmision=carrera, status=True):
            #     print(f"{x.carreraadmision} - {x.carrerapregrado}")

            row_num += 1

            DEBUG and print(f"{counter + 1}.- {carrera} [{carrera.pk}] - {porcentajetotal:.2f} %")
        workbook.close()
    except Exception as ex:
        pass


def reporte_permanencia_estudiantes_poa():
    try:
        import xlsxwriter

        __author__ = 'Unemi'
        ahora = datetime.now()
        time_codigo = ahora.strftime('%Y%m%d%H%M%S')
        name_file = f'000_permanencia_estudiantes_{time_codigo}.xlsx'

        output_folder = f"{BASE_DIR}/media/arreglo_jefferson/{name_file}"
        workbook = xlsxwriter.Workbook(output_folder)
        ws = workbook.add_worksheet('resultados')

        periodo_actual, periodo_pasado = 317, 224

        carreras = Materia.objects.values_list('asignaturamalla__malla__carrera', flat=True).filter(
            asignaturamalla__malla__carrera__coordinacion__in=[1, 2, 3, 4, 5], nivel__periodo=periodo_actual,
            status=True, nivel__status=True, asignaturamalla__status=True,
            asignaturamalla__malla__status=True).distinct()

        fuentecabecera = workbook.add_format({'align': 'center', 'bg_color': 'silver', 'border': 1, 'bold': 1})
        formatoceldacenter = workbook.add_format({'border': 1, 'valign': 'vcenter', 'align': 'center'})
        formatoceldacenter_bold = workbook.add_format({'border': 1, 'valign': 'vcenter', 'align': 'center', 'bold': 1})
        formatoceldaleft = workbook.add_format({'border': 1, 'valign': 'vleft', 'align': 'left'})
        formatoceldaleft_bold = workbook.add_format({'border': 1, 'valign': 'vcenter', 'align': 'left', 'bold': 1})
        fuenteencabezado = workbook.add_format(
            {'align': 'center', 'bg_color': '#1C3247', 'font_color': 'white', 'border': 1, 'font_size': 24, 'bold': 1})

        # ((Número de estudiantes matriculados en la carrera que iniciaron el primer nivel el semestre anterior y se mantienen matriculados en el semestre en curso
        # /
        # Número de estudiantes matriculados en la carrera que iniciaron el primer nivel el semestre anterior)
        # * 100)

        columnas = [(u"#", 10),
                    (u"Carrera", 90),
                    (f"Iniciaron 1ro nivel semestre anterior", 40),
                    (u"Iniciaron 1ro nivel semestre anterior y se mantienen", 40),
                    (f"Porcentaje total", 20)
                    ]
        ws.merge_range(0, 0, 0, len(columnas) - 1, 'UNIVERSIDAD ESTATAL DE MILAGRO', fuenteencabezado)
        ws.merge_range(1, 0, 1, len(columnas) - 1, f'Reporte de permanencia de la carrera', fuenteencabezado)
        row_num, numcolum = 2, 0
        for col_name in columnas:
            ws.write(row_num, numcolum, col_name[0], fuentecabecera)
            ws.set_column(numcolum, numcolum, col_name[1])
            numcolum += 1
        row_num += 1
        for counter, carrera in enumerate(Carrera.objects.filter(id__in=carreras, activa=True, status=True)):
            filtrobase = Q(materia__asignaturamalla__malla__carrera=carrera, retiramateria=False, retiromanual=False,
                           materia__status=True, matricula__inscripcion__persona__real=True, status=True)

            ws.write(row_num, 0, f"{counter + 1}", formatoceldacenter)
            ws.write(row_num, 1, f"{carrera}", formatoceldaleft)

            # Determina si aprobaron -- materiaasignada__estado = 1
            # Número de estudiantes matriculados en la carrera que iniciaron el primer nivel el semestre anterior
            matriculadossemestreanterior = MateriaAsignada.objects.values_list('matricula__inscripcion__persona',
                                                                               flat=True).filter(
                filtrobase & Q(matricula__nivelmalla=1, materia__cerrado=True,
                               materia__nivel__periodo=periodo_pasado)).order_by(
                'matricula__inscripcion__persona__id').distinct('matricula__inscripcion__persona__id')

            filtrobase &= Q(matricula__inscripcion__persona__in=matriculadossemestreanterior)

            # Consultar si matricula__nivelmalla = 2
            # Número de estudiantes matriculados en la carrera que iniciaron el primer nivel el semestre anterior
            matriculadossemestreanterioryencurso = MateriaAsignada.objects.values_list(
                'matricula__inscripcion__persona', flat=True).filter(
                filtrobase & Q(materia__nivel__periodo=periodo_actual)).order_by(
                'matricula__inscripcion__persona__id').distinct('matricula__inscripcion__persona__id')

            a, b = matriculadossemestreanterioryencurso.count(), matriculadossemestreanterior.count()

            porcentajetotal = 0

            try:
                porcentajetotal = ((a / b) * 100)
            except Exception as ex:
                ...

            ws.write(row_num, 2, f"{b}", formatoceldacenter)
            ws.write(row_num, 3, f"{a}", formatoceldacenter)
            ws.write(row_num, 4, f"{porcentajetotal:.2f} %", formatoceldacenter)

            row_num += 1

            DEBUG and print(f"{counter + 1}.- {carrera}[{carrera.pk}] - {porcentajetotal:.2f} %")
        workbook.close()
    except Exception as ex:
        pass


def actualizar_evidencia_internadorotativo():
    try:
        from datetime import date
        from sga.models import EvidenciaActividadDetalleDistributivo, SubTemaAsistencia

        # EvidenciaActividadDetalleDistributivo.objects.filter(id__in=(108316, 108494, 106570, 106638)).update(estadoaprobacion=5)
        for evidencia in EvidenciaActividadDetalleDistributivo.objects.filter(id__in=(106638, 106570)):
            evidencia.estadoaprobacion = 5
            evidencia.save()
            if historial := evidencia.historialaprobacionevidenciaactividad_set.filter(status=True).order_by(
                    '-id').first():
                if not int(historial.estadoaprobacion) == 5:
                    e = HistorialAprobacionEvidenciaActividad(evidencia=evidencia, aprobacionpersona_id=8,
                                                              observacion='Documento legalizado',
                                                              fechaaprobacion=date(2024, 6, 21), estadoaprobacion=5)
                    e.save()

            for anexo in evidencia.evidenciaactividadaudi_set.filter(status=True):
                if not 'media' in anexo.archivo:
                    anexo.archivo = 'media/' + anexo.archivo
                    anexo.save()

    except Exception as ex:
        pass


def arreglo_v3():
    try:
        import random
        from datetime import time
        from sagest.models import LogMarcada
        listasecuencia, listavalores = [1, 2, 3, 4], []
        listavalores.append((1, [time(7, 58), time(7, 59), time(8, 0), time(8, 1)]))
        listavalores.append((2, [time(13, 0), time(13, 3), time(13, 5), time(13, 8)]))
        listavalores.append((3, [time(13, 50), time(13, 53), time(13, 59), time(14, 0)]))
        listavalores.append((4, [time(17, 10), time(17, 13), time(17, 19), time(17, 30)]))

        mes = datetime.now().month - 1 if datetime.now().day < 3 else datetime.now().month
        for persona in Persona.objects.filter(id__in=[37121]):
            for dia in persona.logdia_set.filter(fecha__year=datetime.now().year, fecha__month=mes,
                                                 status=True).order_by('fecha'):
                if marcadas := dia.logmarcada_set.all().order_by('time'):
                    replic = marcadas.first()
                    if replic.time.time() > time(8, 0):
                        t = listavalores[0][1][random.randint(0, 3)]
                        replic.time = replic.time.replace(hour=t.hour, minute=t.minute)
                        replic.save()
                    for m in marcadas:
                        if m.time.time() < time(9, 0):
                            listasecuencia.remove(1)
                        if time(13, 0) <= m.time.time() < time(13, 30):
                            listasecuencia.remove(2)
                        if time(13, 30) <= m.time.time() < time(14, 0):
                            listasecuencia.remove(3)
                        if m.time.time() > time(17, 0):
                            listasecuencia.remove(4)
                    for s in listavalores:
                        if s[0] in listasecuencia:
                            try:
                                t = s[1][random.randint(0, 3)]
                                new = replic.time.replace(hour=t.hour, minute=t.minute)
                                LogMarcada(logdia=dia, time=new, secuencia=s[0], manual=replic.manual,
                                           tipomarcada=replic.tipomarcada, ipmarcada=replic.ipmarcada,
                                           direccion=replic.direccion).save()
                            except Exception as ex:
                                ...
    except Exception as ex:
        pass


def arreglo_v4():
    try:
        for persona in Persona.objects.filter(id__in=[37121]):
            for dia in persona.logdia_set.filter(fecha__year=2024, fecha__month=6,
                                                 fecha__day__in=(3, 4, 5, 6, 7, 10, 14, 18, 21), status=True).order_by(
                    'fecha'):
                queryset = dia.logmarcada_set.filter(status=True).order_by('time')
                for i, m in enumerate(queryset):
                    if i == 1 and len(queryset) == 3:
                        m.delete()
    except Exception as ex:
        pass


def justificar_asistencia_semana7():
    try:
        from sga.models import SubTemaAsistencia, DetalleSilaboSemanalSubtema
        subtema = DetalleSilaboSemanalSubtema.objects.filter(id=2032198)
        if not subtema.subtemaasistencia_set.filter(tema=1039690, status=True):
            subtema = SubTemaAsistencia(tema_id=1039690, subtema=subtema, fecha=date(2024, 5, 13))
            subtema.save()
    except Exception as ex:
        pass


def eliminar_informes_turismo_presencial():
    try:
        import os
        from settings import SITE_STORAGE
        from sga.models import ProfesorDistributivoHoras

        # Eliminar informe
        LISTADO_MESES_ELIMINAR = [4, 5]
        for distributivo in ProfesorDistributivoHoras.objects.filter(periodo=317, carrera=138, activo=True,
                                                                     status=True):
            for mes in LISTADO_MESES_ELIMINAR:
                for informemensual in distributivo.informemensualdocente_set.filter(fechafin__year=2024,
                                                                                    fechafin__month=mes, status=True):
                    generado = 'informemensual_' + str(informemensual.distributivo.id) + '_' + str(
                        informemensual.fechafin.month)
                    firmado = 'informemensual_' + str(informemensual.distributivo.id) + '_' + str(
                        informemensual.fechafin.month) + '_2'
                    folder = os.path.join(SITE_STORAGE, 'media', 'informemensualdocente', '')

                    rutapdf = folder + generado + '.pdf'
                    try:
                        os.path.isfile(rutapdf) and os.remove(rutapdf)
                    except Exception as ex:
                        pass

                    rutapdf = folder + firmado + '.pdf'
                    try:
                        os.path.isfile(rutapdf) and os.remove(rutapdf)
                    except Exception as ex:
                        pass
                    informemensual.delete()
    except Exception as ex:
        pass


def justificar_semana_7():
    import os
    import sys
    from settings import DEBUG, SITE_STORAGE
    from sga.funciones import notificacion2
    from datetime import datetime, timedelta
    import runback.arreglos.arreglo_jefferson
    from sga.models import Silabo, Periodo, Persona, TemaAsistencia, SubTemaAsistencia, Leccion, ProfesorMateria, \
        DetalleSilaboSemanalTema, DetalleSilaboSemanalSubtema, ProfesorDistributivoHoras
    SEMANAS_JUSTIFICADAS = [4]
    periodo = Periodo.objects.get(id=317)
    persona_notifica = Persona.objects.get(id=37121)
    hoy = datetime.now().date()

    # Justificar semana 4
    try:
        lista_afectados = []
        for profesormateria in ProfesorMateria.objects.filter(id=161730).distinct('materia'):
            if silabo := Silabo.objects.filter(profesor=profesormateria.profesor, materia=profesormateria.materia,
                                               status=True).first():
                # Lección al azar por si no se aperturó ninguna entre la semana del silabo
                leccion = Leccion.objects.filter(clase__materia=profesormateria.materia, status=True).order_by(
                    'fecha_creacion').first()
                for ss in silabo.silabosemanal_set.filter(status=True, numsemana__in=SEMANAS_JUSTIFICADAS):
                    if __leccion := Leccion.objects.filter(clase__materia=profesormateria.materia,
                                                           fecha__gte=ss.fechainiciosemana,
                                                           fecha__lte=ss.fechafinciosemana, status=True).order_by(
                            'fecha_creacion').first():
                        leccion = __leccion
                    filtro_silabos_1_y_2 = Q(silabosemanal__silabo=silabo, silabosemanal=ss, status=True)
                    # No se excluyen los temas marca2 porque pueden haber temas sin subtemas marca2
                    for tema in DetalleSilaboSemanalTema.objects.filter(filtro_silabos_1_y_2):
                        fecharegistro = leccion.fecha
                        fi, ff = tema.silabosemanal.fechainiciosemana, tema.silabosemanal.fechafinciosemana + timedelta(
                            1)
                        if not fi <= fecharegistro <= ff:
                            for __date in daterange(fi, ff):
                                if __date.isoweekday() == leccion.clase.dia:
                                    fecharegistro = __date
                        temaasistencia = TemaAsistencia.objects.filter(leccion__clase__materia=profesormateria.materia,
                                                                       tema=tema, status=True).first()
                        if not temaasistencia:
                            temaasistencia = TemaAsistencia(leccion=leccion, tema=tema, fecha=fecharegistro)
                            temaasistencia.save()
                        registro_subtema = 0
                        # Se excluyen los subtemas que ya fueron registrados en la materia
                        __exclude = SubTemaAsistencia.objects.filter(
                            tema__leccion__clase__materia=profesormateria.materia, status=True).values_list('subtema',
                                                                                                            flat=True)
                        for subtema in DetalleSilaboSemanalSubtema.objects.filter(filtro_silabos_1_y_2).exclude(
                                id__in=__exclude):
                            if not SubTemaAsistencia.objects.filter(tema=temaasistencia, subtema=subtema,
                                                                    status=True).values('id').exists():
                                subtemaasistencia = SubTemaAsistencia(tema=temaasistencia, subtema=subtema,
                                                                      fecha=fecharegistro)
                                subtemaasistencia.save()
                                registro_subtema += 1
                        registro_subtema and lista_afectados.append(
                            f"{registro_subtema} - SEMANA {ss.numsemana} - {profesormateria.__str__()}")
                    DEBUG and print(f"{profesormateria}")
        lista_afectados and notificacion2("Resultados arreglo jefferson", f" <br>".join(lista_afectados),
                                          persona_notifica, None, '/notificacion', persona_notifica.pk, 1, 'sga',
                                          Persona)
    except Exception as ex:
        DEBUG and print(ex.__str__())
        linea_error = 'Error on line {}'.format(sys.exc_info()[-1].tb_lineno)
        notificacion2("Resultados arreglo jefferson", f"{linea_error}. {ex.__str__()}", persona_notifica, None,
                      '/notificacion', persona_notifica.pk, 1, 'sga', Persona)

    # Eliminar informes de abril y mayo según lo que pidió el vicerrectorado
    try:
        LISTADO_MESES_ELIMINAR = [4, 5]
        for distributivo in ProfesorDistributivoHoras.objects.filter(periodo=317, carrera=138, activo=True,
                                                                     status=True):
            for mes in LISTADO_MESES_ELIMINAR:
                for informemensual in distributivo.informemensualdocente_set.filter(fechafin__year=2024,
                                                                                    fechafin__month=mes, status=True):
                    generado = 'informemensual_' + str(informemensual.distributivo.id) + '_' + str(
                        informemensual.fechafin.month)
                    firmado = 'informemensual_' + str(informemensual.distributivo.id) + '_' + str(
                        informemensual.fechafin.month) + '_2'
                    folder = os.path.join(SITE_STORAGE, 'media', 'informemensualdocente', '')
                    rutapdf = folder + generado + '.pdf'
                    try:
                        os.path.isfile(rutapdf) and os.remove(rutapdf)
                    except Exception as ex:
                        pass
                    rutapdf = folder + firmado + '.pdf'
                    try:
                        os.path.isfile(rutapdf) and os.remove(rutapdf)
                    except Exception as ex:
                        pass
                    informemensual.delete()
    except Exception as ex:
        pass


def reporte_alexandra():
    try:
        import xlsxwriter
        from django.db.models import F, ExpressionWrapper, DurationField

        __author__ = 'Unemi'
        ahora = datetime.now()
        time_codigo = ahora.strftime('%Y%m%d%H%M%S')
        name_file = f'reporte_permanencia_docente_unemi_{time_codigo}.xlsx'

        output_folder = f"{BASE_DIR}/media/{name_file}"
        workbook = xlsxwriter.Workbook(output_folder)
        for hoja in ['resultados']:
            ws = workbook.add_worksheet(hoja)
            periodo = Periodo.objects.get(id=317)
            fuentecabecera = workbook.add_format({'align': 'center', 'bg_color': 'silver', 'border': 1, 'bold': 1})
            formatoceldacenter = workbook.add_format({'border': 1, 'valign': 'vcenter', 'align': 'center'})
            formatoceldacenter_bold = workbook.add_format(
                {'border': 1, 'valign': 'vcenter', 'align': 'center', 'bold': 1})
            formatoceldaleft = workbook.add_format({'border': 1, 'valign': 'vleft', 'align': 'left'})
            formatoceldaleft_bold = workbook.add_format({'border': 1, 'valign': 'vcenter', 'align': 'left', 'bold': 1})
            fuenteencabezado = workbook.add_format(
                {'align': 'center', 'bg_color': '#1C3247', 'font_color': 'white', 'border': 1, 'font_size': 24,
                 'bold': 1})
            columnas = [(u"#", 10),
                        (u"Profesor".upper(), 30),
                        (u"Cedula".upper(), 20),
                        (u"Titulo 3er nivel ".upper(), 30),
                        (u"Titulo 4to nivel ".upper(), 30),
                        (u"Años Permanencia UNEMI".upper(), 20),
                        ]
            ws.merge_range(0, 0, 0, len(columnas) - 1, 'UNIVERSIDAD ESTATAL DE MILAGRO', fuenteencabezado)
            ws.merge_range(1, 0, 1, len(columnas) - 1, f'Reporte de docentes de 3 y 4 to nivel', fuenteencabezado)
            row_num, numcolum = 3, 0
            for col_name in columnas:
                ws.write(row_num, numcolum, col_name[0], fuentecabecera)
                ws.set_column(numcolum, numcolum, col_name[1])
                numcolum += 1
            row_num += 1
            ws.merge_range(2, 0, 2, len(columnas) - 1, f'{periodo.nombre}', fuenteencabezado)
            # queryset = ProfesorMateria.objects.values_list('profesor__persona', flat=True).filter(materia__nivel__periodo=periodo, materia__asignaturamalla__malla__carrera__coordinacion=4, status=True).order_by('profesor').distinct()
            queryset = ProfesorDistributivoHoras.objects.values_list('profesor__persona', flat=True).filter(
                periodo=periodo, activo=True, profesor__persona__real=True, coordinacion=4, status=True,
                profesor__status=True, profesor__persona__status=True)
            for counter, persona in enumerate(Persona.objects.filter(id__in=[queryset], real=True, status=True)):
                ws.write(row_num, 0, f"{counter + 1}", formatoceldacenter)
                ws.write(row_num, 1, f"{persona}", formatoceldaleft)
                ws.write(row_num, 2, f"{persona.cedula}", formatoceldaleft)

                titulos = ''
                for titulo in persona.mis_titulaciones().filter(titulo__nivel_id=3, status=True):
                    titulos += f'*{titulo.titulo} '

                ws.write(row_num, 3, f"{titulos}", formatoceldaleft)

                titulos = ''
                for titulo in persona.mis_titulaciones().filter(status=True).exclude(titulo__nivel_id=3):
                    titulos += f'*{titulo.titulo} '

                ws.write(row_num, 4, f"{titulos}", formatoceldaleft)

                ingresopersona = persona.ingresopersonal_set.filter(status=True).order_by('fechaingreso').first()
                historialdistr = persona.distributivopersonahistorial_set.filter(status=True).order_by(
                    'fechahistorial').first()
                fechaingreso, aniospermanencia = None, 0
                if ingresopersona:
                    fechaingreso = ingresopersona.fechaingreso
                    if historialdistr and historialdistr.fechahistorial and historialdistr.fechahistorial.date() < fechaingreso:
                        fechaingreso = historialdistr.fechahistorial
                if fechaingreso:
                    aniospermanencia = ahora.year - fechaingreso.year
                ws.write(row_num, 5, f"{aniospermanencia}", formatoceldacenter)
                row_num += 1
        workbook.close()
    except Exception as ex:
        notificacion2("Resultados arreglo jefferson", f"{ex.__str__()}", persona_notifica, None, '/notificacion',
                      persona_notifica.pk, 1, 'sga', Persona)


def validar_bitacoras_sin_revisor():
    from sga.models import Persona
    from sga.funciones import notificacion2
    from investigacion.models import DetalleBitacoraDocente
    persona_notifica = Persona.objects.get(id=37121)
    try:
        afectados = ''
        for e in DetalleBitacoraDocente.objects.filter(estadoaprobacion=1,
                                                       bitacoradocente__subactividad__subactividaddocenteperiodo__criterio=18,
                                                       status=True).distinct():
            b = e.bitacoradocente
            if not b.estadorevision == 3:
                b.estadorevision = 3
                b.save()
            e.estadoaprobacion = 2
            e.save()

            afectados += f"{b.criterio.distributivo.profesor}<br>"
        notificacion2("Resultados arreglo jefferson", f"{afectados}", persona_notifica, None, '/notificacion',
                      persona_notifica.pk, 1, 'sga', Persona)
        print(afectados)
    except Exception as ex:
        notificacion2("Resultados arreglo jefferson", f"{ex.__str__()}", persona_notifica, None, '/notificacion',
                      persona_notifica.pk, 1, 'sga', Persona)


def reporte_solicitud_cupo():
    from sga.funciones import notificacion2
    from matricula.models import DetalleSolicitudReservaCupoMateria
    from sga.models import Persona, Materia, Periodo, AsignaturaMalla
    try:
        import xlsxwriter
        from datetime import datetime, timedelta
        from settings import TIPO_DOCENTE_TEORIA, TIPO_DOCENTE_PRACTICA, SITE_STORAGE

        __author__ = 'Unemi'
        ahora = datetime.now()
        time_codigo = ahora.strftime('%Y%m%d%H%M%S')
        name_file = f'00_reporte_solicitud_cupo{time_codigo}.xlsx'
        output_folder = f"{SITE_STORAGE}/media/{name_file}"
        workbook = xlsxwriter.Workbook(output_folder)
        ws = workbook.add_worksheet('RESULTADOS')

        periodo = Periodo.objects.get(id=336)
        HOST = "https://sga.unemi.edu.ec"
        ePersona = Persona.objects.get(cedula='0604551580')
        eNotificacion = Notificacion(titulo='Reporte solicitud cupos',
                                     cuerpo=f"""SITE_STORAGE: {SITE_STORAGE}""",
                                     destinatario=ePersona,
                                     url=f'{HOST}/media/{name_file}',
                                     prioridad=1,
                                     app_label='sga',
                                     fecha_hora_visible=datetime.now() + timedelta(days=2),
                                     tipo=1,
                                     content_type=None,
                                     object_id=None,
                                     )
        eNotificacion.save()

        justthey = DetalleSolicitudReservaCupoMateria.objects.filter(solicitud__periodomatricula__periodo=periodo,
                                                                     status=True, solicitud__status=True).order_by(
            'asignaturamalla')
        # queryset = AsignaturaMalla.objects.filter(id__in=justthey, malla__vigente=True, status=True)[:10]

        fuentecabecera = workbook.add_format({'align': 'center', 'bg_color': 'silver', 'border': 1, 'bold': 1})
        formatoceldacenter = workbook.add_format({'border': 1, 'valign': 'vcenter', 'align': 'center'})
        formatoceldacenter_bold = workbook.add_format({'border': 1, 'valign': 'vcenter', 'align': 'center', 'bold': 1})
        formatoceldaleft = workbook.add_format({'border': 1, 'valign': 'vleft', 'align': 'left'})
        formatoceldaleft_bold = workbook.add_format({'border': 1, 'valign': 'vcenter', 'align': 'left', 'bold': 1})
        fuenteencabezado = workbook.add_format(
            {'align': 'center', 'bg_color': '#1C3247', 'font_color': 'white', 'border': 1, 'font_size': 24, 'bold': 1})

        columnas = [(u"#", 10),
                    (u"FACULTAD", 40),
                    (f"CARRERA", 60),
                    (u"MALLA", 40),
                    (f"SECCIÓN", 20),
                    (f"NIVEL", 20),
                    (f"PARALELO", 20),
                    (f"ASIGNATURA", 60),
                    (f"TEORICA PRACTICA", 20),
                    (f"TOTAL DE MATRICULADOS", 20),
                    (f"SOLICITUD DE CUPO", 20),
                    (f"DOCENTE", 20),
                    (f"CEDULA", 20),
                    (f"HORAS SEMANALES", 20),
                    (f"MALLA (HORAS PRESENCIALES SEMANALES)", 20),
                    (f"DEDICACIÓN", 20),
                    (f"CATEGORÍA", 20),
                    (f"MODELO EVALUATIVO", 20),
                    (f"ID MATERIA", 20),
                    (f"MODALIDAD IMPARTICIÓN", 20),
                    ]
        ws.merge_range(0, 0, 0, len(columnas) - 1, 'UNIVERSIDAD ESTATAL DE MILAGRO', fuenteencabezado)
        ws.merge_range(1, 0, 1, len(columnas) - 1, f'Reporte de solicitud de cupos por asignatura', fuenteencabezado)
        row_num, numcolum = 2, 0
        for col_name in columnas:
            ws.write(row_num, numcolum, col_name[0], fuentecabecera)
            ws.set_column(numcolum, numcolum, col_name[1])
            numcolum += 1

        row_num += 1
        for i, solicitud in enumerate(justthey, start=1):
            try:
                asignaturamalla = solicitud.asignaturamalla

                teoriapractica, totalmatricula = 'NO', 0
                docente, cedula, dedicacion, categoria = ['', '', '', '']
                if materia := Materia.objects.filter(asignaturamalla=asignaturamalla, nivel__periodo=periodo,
                                                     status=True).first():
                    totalmatricula = materia.materiaasignada_set.values("id").filter(status=True, retiramateria=False,
                                                                                     retiromanual=False).count()
                    if profesor := materia.profesor_principal():
                        docente, cedula, dedicacion, categoria = profesor.persona, profesor.persona.cedula, profesor.dedicacion, profesor.categoria

                solicitudcupos = asignaturamalla.detallesolicitudreservacupomateria_set.filter(
                    solicitud__periodomatricula__periodo=periodo, status=True, solicitud__status=True).count()

                ws.write(row_num, 0, f"{i}", formatoceldacenter)
                ws.write(row_num, 1, f"{asignaturamalla.malla.carrera.coordinacion_set.first()}", formatoceldacenter)
                ws.write(row_num, 2, f"{asignaturamalla.malla.carrera}", formatoceldacenter)
                ws.write(row_num, 3, f"{asignaturamalla.malla}", formatoceldacenter)
                ws.write(row_num, 4, f"{materia.nivel.sesion if materia else ''}", formatoceldacenter)
                ws.write(row_num, 5, f"{asignaturamalla.nivelmalla}", formatoceldacenter)
                ws.write(row_num, 6, f"{materia.paralelo if materia else ''}", formatoceldacenter)
                ws.write(row_num, 7, f"{asignaturamalla.asignatura}", formatoceldacenter)
                ws.write(row_num, 8, f"{'SI' if asignaturamalla.practicas else 'NO'}", formatoceldacenter)
                ws.write(row_num, 9, f"{totalmatricula}", formatoceldacenter)
                ws.write(row_num, 10, f"{'SI' if solicitudcupos else 'NO'}", formatoceldacenter)
                ws.write(row_num, 11, f"{docente}", formatoceldacenter)
                ws.write(row_num, 12, f"{cedula}", formatoceldacenter)
                ws.write(row_num, 13, f"{materia.horassemanales if materia else 0}", formatoceldacenter)
                ws.write(row_num, 14, f"{asignaturamalla.horaspresencialessemanales}", formatoceldacenter)
                ws.write(row_num, 15, f"{dedicacion}", formatoceldacenter)
                ws.write(row_num, 16, f"{categoria}", formatoceldacenter)
                ws.write(row_num, 17, f"{materia.modeloevaluativo if materia else ''}", formatoceldacenter)
                ws.write(row_num, 18, f"{materia.pk if materia else ''}", formatoceldacenter)
                ws.write(row_num, 19, f"{asignaturamalla.malla.carrera.get_modalidad_display()}", formatoceldacenter)

                row_num += 1
            except Exception as ex:
                ...
        workbook.close()
        for e in Persona.objects.filter(id__in=[29898, 23266, 37121]):
            notificacion2("Reporte solicitud de cupos", f"Se ejecutó correctamente...", e, None,
                          f'{HOST}/media/{name_file}', e.pk, 1, 'sga', Persona)
    except Exception as ex:
        e = Persona.objects.get(id=37121)
        notificacion2("Error reporte solicitud de cupos", f"{ex=}", e, None, '/notificacion', e.pk, 1, 'sga', Persona)


def llenar_tablas_local():
    try:
        import pandas as pd
        from sga.models import Periodo, Inscripcion, AsignaturaMalla, Sesion
        from matricula.models import SolicitudReservaCupoMateria, DetalleSolicitudReservaCupoMateria, PeriodoMatricula

        # Lee el archivo Excel
        file_path = '/media/solicitud_cupos_produccion.xlsx'
        df = pd.read_excel(f"{BASE_DIR}/{file_path}")
        for index, row in df.iterrows():
            if index > 0:
                solicitud, created = SolicitudReservaCupoMateria.objects.get_or_create(periodo_id=row[15],
                                                                                       inscripcion_id=row[14],
                                                                                       estado=row[13],
                                                                                       periodomatricula_id=row[16])
                detalle = DetalleSolicitudReservaCupoMateria(solicitud=solicitud, asignaturamalla_id=row[4],
                                                             sesion_id=row[5])
                detalle.save()
        print("Datos importados exitosamente.")
    except Exception as ex:
        pass


def llenar_tabla_requisitos():
    try:
        from inno.models import RequerimientoInternoPractica

        listinsert = ['Videocamara de mano', 'Baterias', 'Adaptador de tarjetas', 'Tarjetas SD',
                      'Adaptador de cargador', 'Cable de alimentación', 'Correa para hombro']
        for l in listinsert:
            if not RequerimientoInternoPractica.objects.filter(detalle=l.upper(), status=True):
                object = RequerimientoInternoPractica(tipo=2, detalle=l.upper(), ubicacion=1)
                object.save()
    except Exception as ex:
        pass


def eliminar_silabosemanal_duplicado():
    from settings import DEBUG
    from sga.funciones import notificacion2
    from sga.models import Silabo, Persona, EvaluacionAprendizajeSilaboSemanal
    detalle = ''
    pObj = Persona.objects.get(cedula='0606274652')
    try:
        for silabo in Silabo.objects.filter(id__in=[52027, 52020]):
            silabossemanales = silabo.silabosemanal_set.filter(status=True)
            if len(silabossemanales) > 15:
                for i, value in enumerate(silabossemanales):
                    try:
                        nextss = silabossemanales[i + 1]
                        if value.numsemana == nextss.numsemana:
                            objetoeliminar = value
                            # Si en la semana actual hay planificacion de recursos
                            if EvaluacionAprendizajeSilaboSemanal.objects.filter(silabosemanal__silabo=value.silabo,
                                                                                 tipoactividadsemanal=1,
                                                                                 silabosemanal=value, status=True,
                                                                                 silabosemanal__status=True):
                                objetoeliminar = nextss
                            objetoeliminar.status = False
                            objetoeliminar.save()

                            detalle += f'Se eliminó {objetoeliminar.pk} - {objetoeliminar.numsemana} <br> '
                    except Exception as ex:
                        continue
        DEBUG and print(detalle)
        notificacion2("Error arreglo eliminar duplicados", f"Ejecutado correctamente. {detalle}", pObj, None,
                      '/notificacion', pObj.pk, 1, 'sga', Persona)
    except Exception as ex:
        DEBUG and print(detalle)
        notificacion2("Error arreglo eliminar duplicados", f"{ex=}", pObj, None, '/notificacion', pObj.pk, 1, 'sga',
                      Persona)


def crear_matricula():
    from datetime import datetime
    from sga.funciones import fechatope
    try:
        if not Matricula.objects.filter(inscripcion=359170, nivel=1725, status=True).exists():
            matricula = Matricula(inscripcion_id=359170,
                                  nivel_id=1725,
                                  pago=False,
                                  iece=False,
                                  becado=False,
                                  porcientobeca=0,
                                  fecha=datetime.now().date(),
                                  hora=datetime.now().time(),
                                  fechatope=fechatope(datetime.now().date()),
                                  termino=True,
                                  fechatermino=datetime.now())
            matricula.save()
            matricula.grupo_socio_economico(1)
            matricula.confirmar_matricula()
            print('Listo...')
    except Exception as ex:
        print(f'{ex}')


def carga_de_candidatos_a_presidentes_actualizado():
    from inno.models import EstudiantesCandidatosaPresidentesdeCurso
    from sga.models import MatriculacionPrimerNivelCarrera, Periodo, AsignaturaMalla, Matricula, MateriaAsignada, \
        RecordAcademico, Notificacion
    from collections import Counter
    from django.db import transaction
    try:
        notarecord = 0.0
        bandera = False
        matriculas = None
        inscripciones = []
        carrera_admision = 0
        contador_registro = 0
        cant_materias_estudiante = 0
        id_configuracion_admision_pregrado = 0
        periodo = Periodo.objects.get(status=True, pk=336)
        id_configuracion_admision_pregrado = MatriculacionPrimerNivelCarrera.objects.values_list('configuracion_id', flat=True).filter(status=True, ejecutoaccion=True).order_by('configuracion_id').last()
        carreras = [133]
        for carrera in carreras:
            with transaction.atomic():
                try:
                    # ELIMINAR LISTADO ANTERIOR PARA REALIZAR ALGUNA MODIFICACION
                    EstudiantesCandidatosaPresidentesdeCurso.objects.filter(status=True, carrera_id=carrera, periodo_id=int(periodo.pk)).update(status=False)
                    niveles_carrera = AsignaturaMalla.objects.values_list('nivelmalla__orden', 'nivelmalla__nombre').filter(status=True, malla__carrera_id=carrera, malla__status=True).order_by('nivelmalla__orden').distinct()

                    for nivel in niveles_carrera:
                        matriculas = Matricula.objects.filter(status=True, nivel__periodo=periodo, inscripcion__carrera_id=carrera, nivelmalla__orden=nivel[0]).order_by('nivelmalla__orden')
                        for matricula in matriculas:
                            mi_malla = matricula.inscripcion.mi_malla()
                            cant_asignatura_malla = AsignaturaMalla.objects.filter(status=True, malla=mi_malla, malla__carrera=matricula.inscripcion.carrera, nivelmalla=matricula.nivelmalla).exclude(itinerario__gt=0).exclude(ejeformativo_id__in=[4, 9, 11, 12]).count()
                            cant_materias_estudiante_total = MateriaAsignada.objects.filter(status=True, matricula=matricula, retiramateria=False).exclude(materia__asignaturamalla__itinerario__gt=0).exclude(materia__asignaturamalla__ejeformativo_id__in=[4, 9, 11, 12]).count()
                            if cant_asignatura_malla == cant_materias_estudiante_total:
                                cant_materias_estudiante_nivel = MateriaAsignada.objects.filter(status=True, matricula=matricula, retiramateria=False, materia__asignaturamalla__nivelmalla=matricula.nivelmalla).exclude(materia__asignaturamalla__itinerario__gt=0).exclude(materia__asignaturamalla__ejeformativo_id__in=[4, 9, 11, 12]).count()
                                if cant_asignatura_malla == cant_materias_estudiante_nivel:
                                    if matricula.nivelmalla.id == 1:
                                        carrera_admision = list(MatriculacionPrimerNivelCarrera.objects.values_list('carreraadmision_id', flat=True).filter(status=True, carrerapregrado_id=carrera, configuracion_id=id_configuracion_admision_pregrado).distinct())[0]
                                        materia_asignada_admision = MateriaAsignada.objects.values_list('notafinal', flat=True).filter(status=True, matricula__inscripcion__persona__cedula=matricula.inscripcion.persona.cedula, matricula__inscripcion__carrera__id=carrera_admision, estado=1)
                                        if materia_asignada_admision:
                                            notarecord = round(sum(materia_asignada_admision) / materia_asignada_admision.count(), 2)
                                    else:
                                        notarecord = matricula.inscripcion.promedio_record()
                                    if es_estudiante_regular(matricula.inscripcion, periodo):
                                        materiasasignadas = MateriaAsignada.objects.filter(status=True, matricula=matricula, retiramateria=False)
                                        bandera = False
                                        if materiasasignadas:
                                            for materiaasignada in materiasasignadas:
                                                if materiaasignada.cantidad_matriculas() > 1:
                                                    bandera = True
                                                    break

                                            if not bandera:
                                                paralelos_estudiante = list(materiasasignadas.values_list('materia__paralelo', flat=True))
                                                contador = Counter(paralelos_estudiante)
                                                paralelo_perteneciente = contador.most_common(1)[0][0]
                                                inscripciones.append(matricula.nivelmalla.nombre)
                                                sexo = 'MUJER' if matricula.inscripcion.persona.es_mujer() else 'HOMBRE'

                                                if not EstudiantesCandidatosaPresidentesdeCurso.objects.filter(periodo_id=periodo.id, status=True, matricula=matricula.id, inscripcion=matricula.inscripcion.pk, carrera_id=carrera).exists():
                                                    presidentes = EstudiantesCandidatosaPresidentesdeCurso(status=True,
                                                                                                           periodo_id=periodo.id,
                                                                                                           periodo_nombre=str(periodo.nombre),
                                                                                                           carrera_id=int(carrera),
                                                                                                           inscripcion=int(matricula.inscripcion.pk),
                                                                                                           matricula=int(matricula.pk),
                                                                                                           cedula=str(matricula.inscripcion.persona.cedula),
                                                                                                           nombres=str(matricula.inscripcion.persona.nombre_completo_inverso()),
                                                                                                           correo=str(matricula.inscripcion.persona.emailinst),
                                                                                                           telefono=str(matricula.inscripcion.persona.telefono),
                                                                                                           sexo=sexo,
                                                                                                           carrera=str(matricula.inscripcion.carrera.nombre),
                                                                                                           orden=int(nivel[0]),
                                                                                                           nivel=str(nivel[1]),
                                                                                                           paralelo=f"{paralelo_perteneciente}",
                                                                                                           promedio_final=str(notarecord))
                                                    presidentes.save()
                                                    contador_registro += 1
                                                    print(str(contador_registro) + ' - ' + str(matricula.inscripcion.persona.cedula) + ' - ' + str(matricula.inscripcion.persona.nombre_completo_inverso()) + ' - ' + str(matricula.inscripcion.carrera.nombre) + ' - ' + str(nivel[1]) + ' - ' + str(paralelo_perteneciente) + ' - ' + str(notarecord))
                except Exception as ex:
                    transaction.set_rollback()
                    raise
    except Exception as ex:
        noti = Notificacion(titulo='Error', cuerpo='Ha ocurrido un error {} - Error en la linea {}'.format(ex, sys.exc_info()[-1].tb_lineno), destinatario_id=29898, url="", prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1), tipo=2, en_proceso=False, error=True)
        noti.save()
        transaction.set_rollback(True)
        print('error: %s' % ex)


def es_estudiante_regular(eInscripcion, ePeriodo):
    ...


def cargar_tabla_requerimeinto_interno():
    import xlwt
    import openpyxl
    from sga.funciones import notificacion2
    from inno.models import RequerimientoInternoPractica
    persona_notifica = Persona.objects.get(id=29898)
    try:
        path_anexo = "media/listado_requerimiento_interno.xlsx"
        miarchivo = openpyxl.load_workbook(path_anexo)
        lista = miarchivo['Sheet1']
        for i, val in enumerate(lista.rows):
            if i > 0:
                tipo, detalle, ubicacion, _ = val
                if not RequerimientoInternoPractica.objects.filter(detalle=f"{detalle.value.strip()}"):
                    RequerimientoInternoPractica(tipo=int(tipo.value), detalle=f"{detalle.value.strip()}",
                                                 ubicacion=int(ubicacion.value)).save()
                    print(detalle)
        n = RequerimientoInternoPractica.objects.values('id').count()
        notificacion2("Resultados arreglo jefferson", f"Se guardaron {n} registros en RequerimientoInternoPractica",
                      persona_notifica, None, '/notificacion', persona_notifica.pk, 1, 'sga', Persona)
    except Exception as ex:
        notificacion2("Resultados arreglo jefferson", f"{ex.__str__()}", persona_notifica, None, '/notificacion',
                      persona_notifica.pk, 1, 'sga', Persona)


def validar_bitacoras_sin_revisor():
    from sga.models import Persona
    from sga.funciones import notificacion2
    from investigacion.models import DetalleBitacoraDocente
    persona_notifica = Persona.objects.get(id=37121)
    try:
        afectados = ''
        for e in DetalleBitacoraDocente.objects.filter(estadoaprobacion=1,
                                                       bitacoradocente__subactividad__subactividaddocenteperiodo__criterio=18,
                                                       status=True).distinct():
            b = e.bitacoradocente
            if not b.estadorevision == 3:
                b.estadorevision = 3
                b.save()
            e.estadoaprobacion = 2
            e.save()

            afectados += f"{b.criterio.distributivo.profesor}<br>"
        notificacion2("Resultados arreglo jefferson", f"{afectados}", persona_notifica, None, '/notificacion',
                      persona_notifica.pk, 1, 'sga', Persona)
        print(afectados)
    except Exception as ex:
        notificacion2("Resultados arreglo jefferson", f"{ex.__str__()}", persona_notifica, None, '/notificacion',
                      persona_notifica.pk, 1, 'sga', Persona)


def reporte_solicitud_cupo():
    from sga.funciones import notificacion2
    from matricula.models import DetalleSolicitudReservaCupoMateria
    from sga.models import Persona, Materia, Periodo, AsignaturaMalla
    try:
        import xlsxwriter
        from datetime import datetime, timedelta
        from settings import TIPO_DOCENTE_TEORIA, TIPO_DOCENTE_PRACTICA, SITE_STORAGE

        __author__ = 'Unemi'
        ahora = datetime.now()
        time_codigo = ahora.strftime('%Y%m%d%H%M%S')
        name_file = f'00_reporte_solicitud_cupo{time_codigo}.xlsx'
        output_folder = f"{SITE_STORAGE}/media/{name_file}"
        workbook = xlsxwriter.Workbook(output_folder)
        ws = workbook.add_worksheet('RESULTADOS')

        periodo = Periodo.objects.get(id=336)
        HOST = "https://sga.unemi.edu.ec"
        ePersona = Persona.objects.get(cedula='0604551580')
        eNotificacion = Notificacion(titulo='Reporte solicitud cupos',
                                     cuerpo=f"""SITE_STORAGE: {SITE_STORAGE}""",
                                     destinatario=ePersona,
                                     url=f'{HOST}/media/{name_file}',
                                     prioridad=1,
                                     app_label='sga',
                                     fecha_hora_visible=datetime.now() + timedelta(days=2),
                                     tipo=1,
                                     content_type=None,
                                     object_id=None,
                                     )
        eNotificacion.save()

        justthey = DetalleSolicitudReservaCupoMateria.objects.filter(solicitud__periodomatricula__periodo=periodo,
                                                                     status=True, solicitud__status=True).order_by(
            'asignaturamalla')
        # queryset = AsignaturaMalla.objects.filter(id__in=justthey, malla__vigente=True, status=True)[:10]

        fuentecabecera = workbook.add_format({'align': 'center', 'bg_color': 'silver', 'border': 1, 'bold': 1})
        formatoceldacenter = workbook.add_format({'border': 1, 'valign': 'vcenter', 'align': 'center'})
        formatoceldacenter_bold = workbook.add_format({'border': 1, 'valign': 'vcenter', 'align': 'center', 'bold': 1})
        formatoceldaleft = workbook.add_format({'border': 1, 'valign': 'vleft', 'align': 'left'})
        formatoceldaleft_bold = workbook.add_format({'border': 1, 'valign': 'vcenter', 'align': 'left', 'bold': 1})
        fuenteencabezado = workbook.add_format(
            {'align': 'center', 'bg_color': '#1C3247', 'font_color': 'white', 'border': 1, 'font_size': 24, 'bold': 1})

        columnas = [(u"#", 10),
                    (u"FACULTAD", 40),
                    (f"CARRERA", 60),
                    (u"MALLA", 40),
                    (f"SECCIÓN", 20),
                    (f"NIVEL", 20),
                    (f"PARALELO", 20),
                    (f"ASIGNATURA", 60),
                    (f"TEORICA PRACTICA", 20),
                    (f"TOTAL DE MATRICULADOS", 20),
                    (f"SOLICITUD DE CUPO", 20),
                    (f"DOCENTE", 20),
                    (f"CEDULA", 20),
                    (f"HORAS SEMANALES", 20),
                    (f"MALLA (HORAS PRESENCIALES SEMANALES)", 20),
                    (f"DEDICACIÓN", 20),
                    (f"CATEGORÍA", 20),
                    (f"MODELO EVALUATIVO", 20),
                    (f"ID MATERIA", 20),
                    (f"MODALIDAD IMPARTICIÓN", 20),
                    ]
        ws.merge_range(0, 0, 0, len(columnas) - 1, 'UNIVERSIDAD ESTATAL DE MILAGRO', fuenteencabezado)
        ws.merge_range(1, 0, 1, len(columnas) - 1, f'Reporte de solicitud de cupos por asignatura', fuenteencabezado)
        row_num, numcolum = 2, 0
        for col_name in columnas:
            ws.write(row_num, numcolum, col_name[0], fuentecabecera)
            ws.set_column(numcolum, numcolum, col_name[1])
            numcolum += 1

        row_num += 1
        for i, solicitud in enumerate(justthey, start=1):
            try:
                asignaturamalla = solicitud.asignaturamalla

                teoriapractica, totalmatricula = 'NO', 0
                docente, cedula, dedicacion, categoria = ['', '', '', '']
                if materia := Materia.objects.filter(asignaturamalla=asignaturamalla, nivel__periodo=periodo,
                                                     status=True).first():
                    totalmatricula = materia.materiaasignada_set.values("id").filter(status=True, retiramateria=False,
                                                                                     retiromanual=False).count()
                    if profesor := materia.profesor_principal():
                        docente, cedula, dedicacion, categoria = profesor.persona, profesor.persona.cedula, profesor.dedicacion, profesor.categoria

                solicitudcupos = asignaturamalla.detallesolicitudreservacupomateria_set.filter(
                    solicitud__periodomatricula__periodo=periodo, status=True, solicitud__status=True).count()

                ws.write(row_num, 0, f"{i}", formatoceldacenter)
                ws.write(row_num, 1, f"{asignaturamalla.malla.carrera.coordinacion_set.first()}", formatoceldacenter)
                ws.write(row_num, 2, f"{asignaturamalla.malla.carrera}", formatoceldacenter)
                ws.write(row_num, 3, f"{asignaturamalla.malla}", formatoceldacenter)
                ws.write(row_num, 4, f"{materia.nivel.sesion if materia else ''}", formatoceldacenter)
                ws.write(row_num, 5, f"{asignaturamalla.nivelmalla}", formatoceldacenter)
                ws.write(row_num, 6, f"{materia.paralelo if materia else ''}", formatoceldacenter)
                ws.write(row_num, 7, f"{asignaturamalla.asignatura}", formatoceldacenter)
                ws.write(row_num, 8, f"{'SI' if asignaturamalla.practicas else 'NO'}", formatoceldacenter)
                ws.write(row_num, 9, f"{totalmatricula}", formatoceldacenter)
                ws.write(row_num, 10, f"{'SI' if solicitudcupos else 'NO'}", formatoceldacenter)
                ws.write(row_num, 11, f"{docente}", formatoceldacenter)
                ws.write(row_num, 12, f"{cedula}", formatoceldacenter)
                ws.write(row_num, 13, f"{materia.horassemanales if materia else 0}", formatoceldacenter)
                ws.write(row_num, 14, f"{asignaturamalla.horaspresencialessemanales}", formatoceldacenter)
                ws.write(row_num, 15, f"{dedicacion}", formatoceldacenter)
                ws.write(row_num, 16, f"{categoria}", formatoceldacenter)
                ws.write(row_num, 17, f"{materia.modeloevaluativo if materia else ''}", formatoceldacenter)
                ws.write(row_num, 18, f"{materia.pk if materia else ''}", formatoceldacenter)
                ws.write(row_num, 19, f"{asignaturamalla.malla.carrera.get_modalidad_display()}", formatoceldacenter)

                row_num += 1
            except Exception as ex:
                ...
        workbook.close()
        for e in Persona.objects.filter(id__in=[29898, 23266, 37121]):
            notificacion2("Reporte solicitud de cupos", f"Se ejecutó correctamente...", e, None,
                          f'{HOST}/media/{name_file}', e.pk, 1, 'sga', Persona)
    except Exception as ex:
        e = Persona.objects.get(id=37121)
        notificacion2("Error reporte solicitud de cupos", f"{ex=}", e, None, '/notificacion', e.pk, 1, 'sga', Persona)


def llenar_tablas_local():
    try:
        import pandas as pd
        from sga.models import Periodo, Inscripcion, AsignaturaMalla, Sesion
        from matricula.models import SolicitudReservaCupoMateria, DetalleSolicitudReservaCupoMateria, PeriodoMatricula

        # Lee el archivo Excel
        file_path = '/media/solicitud_cupos_produccion.xlsx'
        df = pd.read_excel(f"{BASE_DIR}/{file_path}")
        for index, row in df.iterrows():
            if index > 0:
                solicitud, created = SolicitudReservaCupoMateria.objects.get_or_create(periodo_id=row[15],
                                                                                       inscripcion_id=row[14],
                                                                                       estado=row[13],
                                                                                       periodomatricula_id=row[16])
                detalle = DetalleSolicitudReservaCupoMateria(solicitud=solicitud, asignaturamalla_id=row[4],
                                                             sesion_id=row[5])
                detalle.save()
        print("Datos importados exitosamente.")
    except Exception as ex:
        pass


def llenar_tabla_requisitos():
    try:
        from inno.models import RequerimientoInternoPractica

        listinsert = ['Videocamara de mano', 'Baterias', 'Adaptador de tarjetas', 'Tarjetas SD',
                      'Adaptador de cargador', 'Cable de alimentación', 'Correa para hombro']
        for l in listinsert:
            if not RequerimientoInternoPractica.objects.filter(detalle=l.upper(), status=True):
                object = RequerimientoInternoPractica(tipo=2, detalle=l.upper(), ubicacion=1)
                object.save()
    except Exception as ex:
        pass


def eliminar_silabosemanal_duplicado():
    import sys
    from settings import DEBUG
    from sga.funciones import convertir_fecha, notificacion2, daterange
    from datetime import datetime, date, timedelta
    from django.db.models import Q
    from sga.models import Silabo, Persona, EvaluacionAprendizajeSilaboSemanal, Periodo, Notificacion, TemaAsistencia, SubTemaAsistencia, Leccion, ProfesorMateria, DetalleSilaboSemanalTema, DetalleSilaboSemanalSubtema
    detalle = ''
    pObj = Persona.objects.get(id=37121)
    try:
        for silabo in Silabo.objects.filter(id__in=[52026]):
            silabossemanales = silabo.silabosemanal_set.filter(status=True)
            if len(silabossemanales) > 15:
                for i, value in enumerate(silabossemanales):
                    try:
                        nextss = silabossemanales[i + 1]
                        if value.numsemana == nextss.numsemana:
                            objetoeliminar = value
                            # Si en la semana actual hay planificacion de recursos
                            if EvaluacionAprendizajeSilaboSemanal.objects.filter(silabosemanal__silabo=value.silabo,
                                                                                 tipoactividadsemanal=1,
                                                                                 silabosemanal=value, status=True,
                                                                                 silabosemanal__status=True):
                                objetoeliminar = nextss
                            objetoeliminar.status = False
                            objetoeliminar.save()

                            detalle += f'Se eliminó {objetoeliminar.pk} - {objetoeliminar.numsemana} <br> '
                    except Exception as ex:
                        continue
        DEBUG and print(detalle)
        notificacion2("Error arreglo eliminar duplicados", f"Ejecutado correctamente. {detalle}", pObj, None, '/notificacion', pObj.pk, 1, 'sga', Persona)
    except Exception as ex:
        DEBUG and print(detalle)
        notificacion2("Error arreglo eliminar duplicados", f"{ex=}", pObj, None, '/notificacion', pObj.pk, 1, 'sga', Persona)

    id_profesor_materia = [151607, 151621, 161802]
    persona_notifica = Persona.objects.get(id=37121)
    hoy = datetime.now().date()

    if subtema := SubTemaAsistencia.objects.filter(subtema=2043083, fecha=date(2024, 5, 19), status=True).first():
        SubTemaAsistencia(tema=subtema.tema, subtema_id=2043084, fecha=date(2024, 5, 19), usuario_creacion=subtema.usuario_creacion).save()
    try:
        lista_afectados = []
        periodo = Periodo.objects.get(id=317)
        semanasferiado = list(set([f.isocalendar()[1] for f in periodo.diasnolaborable_set.filter(status=True, activo=True).values_list('fecha', flat=True)]))
        # Tema masivo
        for profesormateria in ProfesorMateria.objects.filter(id__in=id_profesor_materia).distinct('materia'):
            filtro_silabosemanal = Q(status=True, fechainiciosemana__lte=hoy)

            if profesormateria.pk in [151607]:
                filtro_silabosemanal &= Q(numsemana=15)

            if profesormateria.pk in [151621]:
                filtro_silabosemanal &= Q(numsemana=14)

            if profesormateria.pk in [161802]:
                filtro_silabosemanal &= Q(numsemana__in=[7,10])

            if silabo := profesormateria.materia.silabo_actual():
                leccion = Leccion.objects.filter(clase__materia=profesormateria.materia, status=True).order_by('fecha_creacion').first()
                for ss in silabo.silabosemanal_set.filter(filtro_silabosemanal):
                    if __leccion := Leccion.objects.filter(clase__materia=profesormateria.materia, fecha__gte=ss.fechainiciosemana, fecha__lte=ss.fechafinciosemana, status=True).order_by('fecha_creacion').first():
                        leccion = __leccion
                    filtro_silabos_1_y_2 = Q(silabosemanal__silabo=silabo, silabosemanal=ss, status=True)

                    # No se excluyen los temas porque pueden haber temas sin subtemas marca2
                    for tema in DetalleSilaboSemanalTema.objects.filter(filtro_silabos_1_y_2):
                        fecharegistro = leccion.fecha
                        fi, ff = tema.silabosemanal.fechainiciosemana, tema.silabosemanal.fechafinciosemana + timedelta(weeks=1)
                        if fi.isocalendar()[1] in semanasferiado:
                            fi += timedelta(weeks=1)

                        if not fi <= fecharegistro <= ff:
                            for __date in daterange(fi, ff):
                                if __date.isoweekday() == leccion.clase.dia:
                                    fecharegistro = __date
                                    break

                        if temaasistencia := TemaAsistencia.objects.filter(leccion__clase__materia=profesormateria.materia, tema=tema, status=True).first():
                            if temaasistencia.fecha >= ff:
                                temaasistencia.fecha = fecharegistro
                                temaasistencia.save()
                        else:
                            temaasistencia = TemaAsistencia(leccion=leccion, tema=tema, fecha=fecharegistro)
                            temaasistencia.save()

                        registro_subtema = 0

                        # Se excluyen los subtemas que ya fueron registrados en la materia
                        __exclude = SubTemaAsistencia.objects.filter(fecha__lte=ff, tema__leccion__clase__materia=profesormateria.materia, status=True).values_list('subtema', flat=True)
                        for subtema in DetalleSilaboSemanalSubtema.objects.filter(filtro_silabos_1_y_2).exclude(id__in=__exclude):
                            if subtemaasistencia := SubTemaAsistencia.objects.filter(tema=temaasistencia, subtema=subtema, status=True).order_by('fecha').first():
                                if subtemaasistencia.fecha > ff:
                                    subtemaasistencia.fecha = fecharegistro
                                    subtemaasistencia.save()
                                    registro_subtema += 1
                            else:
                                subtemaasistencia = SubTemaAsistencia(tema=temaasistencia, subtema=subtema, fecha=fecharegistro)
                                subtemaasistencia.save()
                                registro_subtema += 1
                        registro_subtema and lista_afectados.append(f"{registro_subtema} - SEMANA {ss.numsemana} - {profesormateria} - fecha registro {fecharegistro}")
                    DEBUG and print(f"{profesormateria}")

        lista_afectados and notificacion2("Resultados arreglo", f" <br>".join(lista_afectados), persona_notifica, None, '/notificacion', persona_notifica.pk, 1, 'sga', Persona)
    except Exception as ex:
        DEBUG and print(ex.__str__())
        linea_error = 'Error on line {}'.format(sys.exc_info()[-1].tb_lineno)
        notificacion2("Resultados arreglo jefferson", f"{linea_error}. {ex.__str__()}", persona_notifica, None, '/notificacion', persona_notifica.pk, 1, 'sga', Persona)


def crear_matricula():
    from datetime import datetime
    from sga.funciones import fechatope
    try:
        if not Matricula.objects.filter(inscripcion=359170, nivel=1725, status=True).exists():
            matricula = Matricula(inscripcion_id=359170,
                                  nivel_id=1725,
                                  pago=False,
                                  iece=False,
                                  becado=False,
                                  porcientobeca=0,
                                  fecha=datetime.now().date(),
                                  hora=datetime.now().time(),
                                  fechatope=fechatope(datetime.now().date()),
                                  termino=True,
                                  fechatermino=datetime.now())
            matricula.save()
            matricula.grupo_socio_economico(1)
            matricula.confirmar_matricula()
            print('Listo...')
    except Exception as ex:
        print(f'{ex}')


def carga_de_candidatos_a_presidentes_actualizado():
    from inno.models import EstudiantesCandidatosaPresidentesdeCurso
    from sga.models import MatriculacionPrimerNivelCarrera, Periodo, AsignaturaMalla, Matricula, MateriaAsignada, \
        RecordAcademico, Notificacion
    from collections import Counter
    from django.db import transaction
    try:
        notarecord = 0.0
        bandera = False
        matriculas = None
        inscripciones = []
        carrera_admision = 0
        contador_registro = 0
        cant_materias_estudiante = 0
        id_configuracion_admision_pregrado = 0
        periodo = Periodo.objects.get(status=True, pk=336)
        id_configuracion_admision_pregrado = MatriculacionPrimerNivelCarrera.objects.values_list('configuracion_id', flat=True).filter(status=True, ejecutoaccion=True).order_by('configuracion_id').last()
        carreras = [133]
        for carrera in carreras:
            with transaction.atomic():
                try:
                    # ELIMINAR LISTADO ANTERIOR PARA REALIZAR ALGUNA MODIFICACION
                    EstudiantesCandidatosaPresidentesdeCurso.objects.filter(status=True, carrera_id=carrera, periodo_id=int(periodo.pk)).update(status=False)
                    niveles_carrera = AsignaturaMalla.objects.values_list('nivelmalla__orden', 'nivelmalla__nombre').filter(status=True, malla__carrera_id=carrera, malla__status=True).order_by('nivelmalla__orden').distinct()

                    for nivel in niveles_carrera:
                        matriculas = Matricula.objects.filter(status=True, nivel__periodo=periodo, inscripcion__carrera_id=carrera, nivelmalla__orden=nivel[0]).order_by('nivelmalla__orden')
                        for matricula in matriculas:
                            mi_malla = matricula.inscripcion.mi_malla()
                            cant_asignatura_malla = AsignaturaMalla.objects.filter(status=True, malla=mi_malla, malla__carrera=matricula.inscripcion.carrera, nivelmalla=matricula.nivelmalla).exclude(itinerario__gt=0).exclude(ejeformativo_id__in=[4, 9, 11, 12]).count()
                            cant_materias_estudiante_total = MateriaAsignada.objects.filter(status=True, matricula=matricula, retiramateria=False).exclude(materia__asignaturamalla__itinerario__gt=0).exclude(materia__asignaturamalla__ejeformativo_id__in=[4, 9, 11, 12]).count()
                            if cant_asignatura_malla == cant_materias_estudiante_total:
                                cant_materias_estudiante_nivel = MateriaAsignada.objects.filter(status=True, matricula=matricula, retiramateria=False, materia__asignaturamalla__nivelmalla=matricula.nivelmalla).exclude(materia__asignaturamalla__itinerario__gt=0).exclude(materia__asignaturamalla__ejeformativo_id__in=[4, 9, 11, 12]).count()
                                if cant_asignatura_malla == cant_materias_estudiante_nivel:
                                    if matricula.nivelmalla.id == 1:
                                        carrera_admision = list(MatriculacionPrimerNivelCarrera.objects.values_list('carreraadmision_id', flat=True).filter(status=True, carrerapregrado_id=carrera, configuracion_id=id_configuracion_admision_pregrado).distinct())[0]
                                        materia_asignada_admision = MateriaAsignada.objects.values_list('notafinal', flat=True).filter(status=True, matricula__inscripcion__persona__cedula=matricula.inscripcion.persona.cedula, matricula__inscripcion__carrera__id=carrera_admision, estado=1)
                                        if materia_asignada_admision:
                                            notarecord = round(sum(materia_asignada_admision) / materia_asignada_admision.count(), 2)
                                    else:
                                        notarecord = matricula.inscripcion.promedio_record()
                                    if es_estudiante_regular(matricula.inscripcion, periodo):
                                        materiasasignadas = MateriaAsignada.objects.filter(status=True, matricula=matricula, retiramateria=False)
                                        bandera = False
                                        if materiasasignadas:
                                            for materiaasignada in materiasasignadas:
                                                if materiaasignada.cantidad_matriculas() > 1:
                                                    bandera = True
                                                    break

                                            if not bandera:
                                                paralelos_estudiante = list(materiasasignadas.values_list('materia__paralelo', flat=True))
                                                contador = Counter(paralelos_estudiante)
                                                paralelo_perteneciente = contador.most_common(1)[0][0]
                                                inscripciones.append(matricula.nivelmalla.nombre)
                                                sexo = 'MUJER' if matricula.inscripcion.persona.es_mujer() else 'HOMBRE'

                                                if not EstudiantesCandidatosaPresidentesdeCurso.objects.filter(periodo_id=periodo.id, status=True, matricula=matricula.id, inscripcion=matricula.inscripcion.pk, carrera_id=carrera).exists():
                                                    presidentes = EstudiantesCandidatosaPresidentesdeCurso(status=True,
                                                                                                           periodo_id=periodo.id,
                                                                                                           periodo_nombre=str(periodo.nombre),
                                                                                                           carrera_id=int(carrera),
                                                                                                           inscripcion=int(matricula.inscripcion.pk),
                                                                                                           matricula=int(matricula.pk),
                                                                                                           cedula=str(matricula.inscripcion.persona.cedula),
                                                                                                           nombres=str(matricula.inscripcion.persona.nombre_completo_inverso()),
                                                                                                           correo=str(matricula.inscripcion.persona.emailinst),
                                                                                                           telefono=str(matricula.inscripcion.persona.telefono),
                                                                                                           sexo=sexo,
                                                                                                           carrera=str(matricula.inscripcion.carrera.nombre),
                                                                                                           orden=int(nivel[0]),
                                                                                                           nivel=str(nivel[1]),
                                                                                                           paralelo=f"{paralelo_perteneciente}",
                                                                                                           promedio_final=str(notarecord))
                                                    presidentes.save()
                                                    contador_registro += 1
                                                    print(str(contador_registro) + ' - ' + str(matricula.inscripcion.persona.cedula) + ' - ' + str(matricula.inscripcion.persona.nombre_completo_inverso()) + ' - ' + str(matricula.inscripcion.carrera.nombre) + ' - ' + str(nivel[1]) + ' - ' + str(paralelo_perteneciente) + ' - ' + str(notarecord))
                except Exception as ex:
                    transaction.set_rollback()
                    raise
    except Exception as ex:
        noti = Notificacion(titulo='Error', cuerpo='Ha ocurrido un error {} - Error en la linea {}'.format(ex, sys.exc_info()[-1].tb_lineno), destinatario_id=29898, url="", prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1), tipo=2, en_proceso=False, error=True)
        noti.save()
        transaction.set_rollback(True)
        print('error: %s' % ex)


def es_estudiante_regular(eInscripcion, ePeriodo):
    try:

        eInscripcion = Inscripcion.objects.get(id=eInscripcion) if f"{eInscripcion}".isdigit() else eInscripcion
        ePeriodo = Periodo.objects.get(id=ePeriodo) if f"{ePeriodo}".isdigit() else ePeriodo
        eMatriculaAnterior = Matricula.objects.filter(status=True, nivel__periodo=ePeriodo, inscripcion=eInscripcion).first()
        regular, nivel_real = True, 1
        if eMatriculaAnterior:
            eInscripcion_mat = eMatriculaAnterior.inscripcion
            nivel_real = eMatriculaAnterior.nivelmalla.orden
        else:
            eInscripcion_mat = eInscripcion
            nivel_real = eInscripcion_mat.mi_nivel().nivel.orden

        EJE_FORMATIVO_VINCULACION, EJE_FORMATIVO_PRACTICAS = 11, 9
        EXCLUDE_EJE_FORMATIVO = [EJE_FORMATIVO_PRACTICAS, EJE_FORMATIVO_VINCULACION]
        eInscripcionMalla_mat = eInscripcion_mat.malla_inscripcion()
        eMalla_mat = eInscripcionMalla_mat.malla
        itinerario = []
        itinerario.append(0)
        if eInscripcionMalla_mat.malla.modalidad.id == 3:
            eAsignaturasMalla = eMalla_mat.asignaturamalla_set.select_related().filter(status=True).exclude(Q(nivelmalla_id=0) | Q(opcional=True)).order_by('nivelmalla', 'ejeformativo').filter(itinerario__in=itinerario)
        else:
            eAsignaturasMalla = eMalla_mat.asignaturamalla_set.select_related().filter(status=True).exclude(Q(nivelmalla_id=0) | Q(opcional=True) | Q(ejeformativo_id__in=EXCLUDE_EJE_FORMATIVO)).order_by('nivelmalla', 'ejeformativo').filter(itinerario__in=itinerario)

        if eInscripcion_mat.itinerario:
            if eInscripcion_mat.itinerario > 0:
                itinerario.append(eInscripcion_mat.itinerario)
                eAsignaturasMalla = eAsignaturasMalla.filter(itinerario__in=itinerario)

        eModulosMalla = ModuloMalla.objects.filter(malla=eInscripcionMalla_mat.malla)
        eNivelMallaAprobadoActual = eInscripcion_mat.recordacademico_set.filter(asignaturamalla__in=eAsignaturasMalla, status=True, aprobada=True).values_list('asignaturamalla__nivelmalla__orden', flat=True).order_by('-asignaturamalla__nivelmalla__orden').first()
        if eNivelMallaAprobadoActual:
            eAsignaturasMallaquedebeaprobar = eAsignaturasMalla.filter(nivelmalla__orden__lte=eNivelMallaAprobadoActual).exclude(Q(asignatura__id__in=eModulosMalla.values_list('asignatura__id', flat=True)) | Q(asignatura__in=AsignaturaMalla.objects.values_list("asignatura__id", flat=True).filter(malla=eMalla_mat)))
            eAsignaturasRecord = eInscripcion_mat.recordacademico_set.filter(asignaturamalla__in=eAsignaturasMalla, status=True)

            if eAsignaturasRecord.values("id").filter(aprobada=False).exists():
                regular = False

            elif eAsignaturasRecord.filter(asignaturamalla__in=eAsignaturasMallaquedebeaprobar, aprobada=True, status=True).count() < eAsignaturasMallaquedebeaprobar.values("id").count():
                regular = False

            elif eAsignaturasRecord.filter(asignaturamalla__in=eAsignaturasMalla.filter(nivelmalla__orden__gte=eNivelMallaAprobadoActual + 1).exclude(Q(asignatura__id__in=eModulosMalla.values_list('asignatura__id', flat=True)) | Q(asignatura__in=AsignaturaMalla.objects.values_list("asignatura__id", flat=True).filter(malla=eMalla_mat))), aprobada=True, status=True).exists():
                regular = False

            if eInscripcionMalla_mat.malla.modalidad.id == 3:
                eAsignaturasMallaHastaNivelActualTodas = eMalla_mat.asignaturamalla_set.select_related().filter(status=True, itinerario__in=itinerario, nivelmalla__orden__lte=eNivelMallaAprobadoActual).exclude(Q(nivelmalla_id=NIVEL_MALLA_CERO) | Q(opcional=True))
            else:
                eAsignaturasMallaHastaNivelActualTodas = eMalla_mat.asignaturamalla_set.select_related().filter(status=True, itinerario__in=itinerario, nivelmalla__orden__lte=eNivelMallaAprobadoActual).exclude(Q(nivelmalla_id=NIVEL_MALLA_CERO) | Q(opcional=True) | Q(ejeformativo_id__in=EXCLUDE_EJE_FORMATIVO))

            asignaturas_sin_aprobar = eAsignaturasMallaHastaNivelActualTodas.exclude(id__in=eAsignaturasRecord.values_list("asignaturamalla_id", flat=True))
            asignatura_esp = eAsignaturasMallaHastaNivelActualTodas.exclude(id__in=eAsignaturasRecord.values_list("asignaturamalla_id", flat=True)).filter(id__in=list(map(int, variable_valor('ASIGNATURAS_ESPECIALES'))))

            if asignaturas_sin_aprobar.exists():
                regular = False
            elif eAsignaturasRecord.values("id").filter(aprobada=False).exists():
                regular = False

            nivel_real = nivel_real + 1 if regular else nivel_real

            if regular and eMatriculaAnterior:
                niv_record = eNivelMallaAprobadoActual
                if niv_record:
                    nivel_real = niv_record + 1

        tienerecordhomologadas = eInscripcion.recordacademico_set.values('id').filter(homologada=True, status=True).exists()
        if not eMatriculaAnterior and not tienerecordhomologadas:
            regular, nivel_real = True, 1

        DEBUG and print(f"Es regular: {regular}")
        return regular, nivel_real
    except Exception as ex:
        return False, ex.__str__()


def carga_de_candidatos_a_presidentes_actualizado(periodo, carrera):
    from inno.models import EstudiantesCandidatosaPresidentesdeCurso
    from sga.models import MatriculacionPrimerNivelCarrera, Periodo, AsignaturaMalla, Matricula, MateriaAsignada, RecordAcademico, Notificacion
    from collections import Counter
    from django.db import transaction
    from datetime import datetime, timedelta
    try:
        notarecord = 0.0
        inscripciones = []
        bandera, matriculas = False, None
        carrera_admision, contador_registro = 0, 0
        id_configuracion_admision_pregrado, cant_materias_estudiante = 0, 0
        id_configuracion_admision_pregrado = MatriculacionPrimerNivelCarrera.objects.values_list('configuracion_id', flat=True).filter(status=True, ejecutoaccion=True).order_by('configuracion_id').last()
        with transaction.atomic():
            try:
                # ELIMINAR LISTADO ANTERIOR PARA REALIZAR ALGUNA MODIFICACION
                EstudiantesCandidatosaPresidentesdeCurso.objects.filter(status=True, carrera_id=carrera.pk, periodo_id=int(periodo.pk)).delete()
                niveles_carrera = AsignaturaMalla.objects.values_list('nivelmalla__orden', 'nivelmalla__nombre').filter(status=True, malla__carrera_id=carrera, malla__status=True).order_by('nivelmalla__orden').distinct()
                for nivel in niveles_carrera:
                    matriculas = Matricula.objects.filter(status=True, nivel__periodo=periodo, inscripcion__carrera_id=carrera, nivelmalla__orden=nivel[0]).order_by('nivelmalla__orden')
                    for i, matricula in enumerate(matriculas, start=1):
                        mi_malla = matricula.inscripcion.mi_malla()
                        cant_asignatura_malla = AsignaturaMalla.objects.filter(status=True, malla=mi_malla, malla__carrera=matricula.inscripcion.carrera, nivelmalla=matricula.nivelmalla).exclude(itinerario__gt=0).exclude(ejeformativo_id__in=[4, 9, 11, 12]).count()
                        cant_materias_estudiante_total = MateriaAsignada.objects.filter(status=True, matricula=matricula, retiramateria=False).exclude(materia__asignaturamalla__itinerario__gt=0).exclude(materia__asignaturamalla__ejeformativo_id__in=[4, 9, 11, 12]).count()
                        if cant_asignatura_malla == cant_materias_estudiante_total:
                            cant_materias_estudiante_nivel = MateriaAsignada.objects.filter(status=True, matricula=matricula, retiramateria=False, materia__asignaturamalla__nivelmalla=matricula.nivelmalla).exclude(materia__asignaturamalla__itinerario__gt=0).exclude(materia__asignaturamalla__ejeformativo_id__in=[4, 9, 11, 12]).count()
                            if cant_asignatura_malla == cant_materias_estudiante_nivel:
                                if matricula.nivelmalla.id == 1:
                                    carrera_admision = list(MatriculacionPrimerNivelCarrera.objects.values_list('carreraadmision_id', flat=True).filter(status=True, carrerapregrado_id=carrera, configuracion_id=id_configuracion_admision_pregrado).distinct())[0]
                                    materia_asignada_admision = MateriaAsignada.objects.values_list('notafinal', flat=True).filter(status=True, matricula__inscripcion__persona__cedula=matricula.inscripcion.persona.cedula, matricula__inscripcion__carrera__id=carrera_admision, estado=1)
                                    if materia_asignada_admision:
                                        notarecord = round(sum(materia_asignada_admision) / materia_asignada_admision.count(), 2)
                                else:
                                    notarecord = matricula.inscripcion.promedio_record()

                                try:
                                    eInscripcion = Inscripcion.objects.get(id=eInscripcion) if f"{eInscripcion}".isdigit() else eInscripcion
                                    ePeriodo = Periodo.objects.get(id=ePeriodo) if f"{ePeriodo}".isdigit() else ePeriodo
                                    eMatriculaAnterior = Matricula.objects.filter(status=True, nivel__periodo=ePeriodo, inscripcion=eInscripcion).first()
                                    regular, nivel_real = True, 1
                                    if eMatriculaAnterior:
                                        eInscripcion_mat = eMatriculaAnterior.inscripcion
                                        nivel_real = eMatriculaAnterior.nivelmalla.orden
                                    else:
                                        eInscripcion_mat = eInscripcion
                                        nivel_real = eInscripcion_mat.mi_nivel().nivel.orden
                                    EJE_FORMATIVO_VINCULACION, EJE_FORMATIVO_PRACTICAS = 11, 9
                                    EXCLUDE_EJE_FORMATIVO = [EJE_FORMATIVO_PRACTICAS, EJE_FORMATIVO_VINCULACION]
                                    eInscripcionMalla_mat = eInscripcion_mat.malla_inscripcion()
                                    eMalla_mat = eInscripcionMalla_mat.malla
                                    itinerario = []
                                    itinerario.append(0)
                                    if eInscripcionMalla_mat.malla.modalidad.id == 3:
                                        eAsignaturasMalla = eMalla_mat.asignaturamalla_set.select_related().filter(status=True).exclude(Q(nivelmalla_id=0) | Q(opcional=True)).order_by('nivelmalla', 'ejeformativo').filter(itinerario__in=itinerario)
                                    else:
                                        eAsignaturasMalla = eMalla_mat.asignaturamalla_set.select_related().filter(status=True).exclude(Q(nivelmalla_id=0) | Q(opcional=True) | Q(ejeformativo_id__in=EXCLUDE_EJE_FORMATIVO)).order_by('nivelmalla', 'ejeformativo').filter(itinerario__in=itinerario)
                                    if eInscripcion_mat.itinerario:
                                        if eInscripcion_mat.itinerario > 0:
                                            itinerario.append(eInscripcion_mat.itinerario)
                                            eAsignaturasMalla = eAsignaturasMalla.filter(itinerario__in=itinerario)
                                    eModulosMalla = ModuloMalla.objects.filter(malla=eInscripcionMalla_mat.malla)
                                    eNivelMallaAprobadoActual = eInscripcion_mat.recordacademico_set.filter(asignaturamalla__in=eAsignaturasMalla, status=True, aprobada=True).values_list('asignaturamalla__nivelmalla__orden', flat=True).order_by('-asignaturamalla__nivelmalla__orden').first()
                                    if eNivelMallaAprobadoActual:
                                        eAsignaturasMallaquedebeaprobar = eAsignaturasMalla.filter(nivelmalla__orden__lte=eNivelMallaAprobadoActual).exclude(Q(asignatura__id__in=eModulosMalla.values_list('asignatura__id', flat=True)) | Q(asignatura__in=AsignaturaMalla.objects.values_list("asignatura__id", flat=True).filter(malla=eMalla_mat)))
                                        eAsignaturasRecord = eInscripcion_mat.recordacademico_set.filter(asignaturamalla__in=eAsignaturasMalla, status=True)
                                        if eAsignaturasRecord.values("id").filter(aprobada=False).exists():
                                            regular = False
                                        elif eAsignaturasRecord.filter(asignaturamalla__in=eAsignaturasMallaquedebeaprobar, aprobada=True, status=True).count() < eAsignaturasMallaquedebeaprobar.values("id").count():
                                            regular = False
                                        elif eAsignaturasRecord.filter(asignaturamalla__in=eAsignaturasMalla.filter(nivelmalla__orden__gte=eNivelMallaAprobadoActual + 1).exclude(Q(asignatura__id__in=eModulosMalla.values_list('asignatura__id', flat=True)) | Q(asignatura__in=AsignaturaMalla.objects.values_list("asignatura__id", flat=True).filter(malla=eMalla_mat))), aprobada=True, status=True).exists():
                                            regular = False
                                        if eInscripcionMalla_mat.malla.modalidad.id == 3:
                                            eAsignaturasMallaHastaNivelActualTodas = eMalla_mat.asignaturamalla_set.select_related().filter(status=True, itinerario__in=itinerario, nivelmalla__orden__lte=eNivelMallaAprobadoActual).exclude(Q(nivelmalla_id=NIVEL_MALLA_CERO) | Q(opcional=True))
                                        else:
                                            eAsignaturasMallaHastaNivelActualTodas = eMalla_mat.asignaturamalla_set.select_related().filter(status=True, itinerario__in=itinerario, nivelmalla__orden__lte=eNivelMallaAprobadoActual).exclude(Q(nivelmalla_id=NIVEL_MALLA_CERO) | Q(opcional=True) | Q(ejeformativo_id__in=EXCLUDE_EJE_FORMATIVO))
                                        asignaturas_sin_aprobar = eAsignaturasMallaHastaNivelActualTodas.exclude(id__in=eAsignaturasRecord.values_list("asignaturamalla_id", flat=True))
                                        asignatura_esp = eAsignaturasMallaHastaNivelActualTodas.exclude(id__in=eAsignaturasRecord.values_list("asignaturamalla_id", flat=True)).filter(id__in=list(map(int, variable_valor('ASIGNATURAS_ESPECIALES'))))
                                        if asignaturas_sin_aprobar.exists():
                                            regular = False
                                        elif eAsignaturasRecord.values("id").filter(aprobada=False).exists():
                                            regular = False
                                        nivel_real = nivel_real + 1 if regular else nivel_real
                                        if regular and eMatriculaAnterior:
                                            niv_record = eNivelMallaAprobadoActual
                                            if niv_record:
                                                nivel_real = niv_record + 1
                                    tienerecordhomologadas = eInscripcion.recordacademico_set.values('id').filter(homologada=True, status=True).exists()
                                    if not eMatriculaAnterior and not tienerecordhomologadas:
                                        regular, nivel_real = True, 1

                                    es_estudiante_regular = regular
                                except Exception as ex:
                                    es_estudiante_regular = False


                                if es_estudiante_regular:
                                    materiasasignadas = MateriaAsignada.objects.filter(status=True, matricula=matricula, retiramateria=False)
                                    bandera = False
                                    if materiasasignadas:
                                        for materiaasignada in materiasasignadas:
                                            if materiaasignada.cantidad_matriculas() > 1:
                                                bandera = True
                                                break
                                        if not bandera:
                                            paralelos_estudiante = list(materiasasignadas.values_list('materia__paralelo', flat=True))
                                            contador = Counter(paralelos_estudiante)
                                            paralelo_perteneciente = contador.most_common(1)[0][0]
                                            inscripciones.append(matricula.nivelmalla.nombre)
                                            sexo = 'MUJER' if matricula.inscripcion.persona.es_mujer() else 'HOMBRE'
                                            if not EstudiantesCandidatosaPresidentesdeCurso.objects.filter(periodo_id=periodo.id, status=True, matricula=matricula.id, inscripcion=matricula.inscripcion.pk, carrera_id=carrera.pk).exists():
                                                presidentes = EstudiantesCandidatosaPresidentesdeCurso(status=True,
                                                                                                       periodo_id=periodo.id,
                                                                                                       periodo_nombre=periodo.nombre,
                                                                                                       carrera_id=carrera.pk,
                                                                                                       inscripcion=matricula.inscripcion.pk,
                                                                                                       matricula=matricula.pk,
                                                                                                       cedula=matricula.inscripcion.persona.cedula,
                                                                                                       nombres=matricula.inscripcion.persona.nombre_completo_inverso(),
                                                                                                       correo=matricula.inscripcion.persona.emailinst,
                                                                                                       telefono=matricula.inscripcion.persona.telefono,
                                                                                                       sexo=sexo,
                                                                                                       carrera=matricula.inscripcion.carrera.nombre,
                                                                                                       orden=int(nivel[0]),
                                                                                                       nivel=str(nivel[1]),
                                                                                                       paralelo=f"{paralelo_perteneciente}",
                                                                                                       promedio_final=str(notarecord))
                                                presidentes.save()
                                                contador_registro += 1
                                                DEBUG and print(f"{i}.- {matricula.inscripcion.persona.nombre_completo_inverso()}")
            except Exception as ex:
                transaction.set_rollback()
                raise
            noti = Notificacion(titulo=f'{carrera}', cuerpo=f'{carrera} ejecutada correctamente...', destinatario_id=37121, url="", prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1), tipo=2, en_proceso=False, error=True)
            noti.save()
    except Exception as ex:
        import sys
        noti = Notificacion(titulo='Error', cuerpo='Ha ocurrido un error {} - Error en la linea {}'.format(ex, sys.exc_info()[-1].tb_lineno), destinatario_id=37121, url="", prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1), tipo=2, en_proceso=False, error=True)
        noti.save()
        transaction.set_rollback(True)
        raise


def carga_de_candidatos_a_presidentes_actualizado_v2(periodo, carrera):
    from inno.models import EstudiantesCandidatosaPresidentesdeCurso
    from sga.models import MatriculacionPrimerNivelCarrera, Periodo, AsignaturaMalla, Matricula, MateriaAsignada, RecordAcademico, Notificacion
    from collections import Counter
    from django.db import transaction
    from datetime import datetime, timedelta
    try:
        notarecord = 0.0
        inscripciones = []
        bandera, matriculas = False, None
        carrera_admision, contador_registro = 0, 0
        id_configuracion_admision_pregrado, cant_materias_estudiante = 0, 0
        id_configuracion_admision_pregrado = MatriculacionPrimerNivelCarrera.objects.values_list('configuracion_id', flat=True).filter(status=True, ejecutoaccion=True).order_by('configuracion_id').last()
        with transaction.atomic():
            try:
                # ELIMINAR LISTADO ANTERIOR PARA REALIZAR ALGUNA MODIFICACION
                EstudiantesCandidatosaPresidentesdeCurso.objects.filter(status=True, carrera_id=carrera.pk, periodo_id=periodo.pk).delete()
                niveles_carrera = AsignaturaMalla.objects.values_list('nivelmalla__orden', 'nivelmalla__nombre').filter(status=True, malla__carrera_id=carrera, malla__status=True).order_by('nivelmalla__orden').distinct()
                for nivel in niveles_carrera:
                    matriculas = Matricula.objects.filter(status=True, nivel__periodo=periodo, inscripcion__carrera_id=carrera, nivelmalla__orden=nivel[0]).order_by('nivelmalla__orden')
                    for i, matricula in enumerate(matriculas, start=1):
                        mi_malla = matricula.inscripcion.mi_malla()
                        cant_asignatura_malla = AsignaturaMalla.objects.filter(status=True, malla=mi_malla, malla__carrera=matricula.inscripcion.carrera, nivelmalla=matricula.nivelmalla).exclude(itinerario__gt=0).exclude(ejeformativo_id__in=[4, 9, 11, 12]).count()
                        cant_materias_estudiante_total = MateriaAsignada.objects.filter(status=True, matricula=matricula, retiramateria=False).exclude(materia__asignaturamalla__itinerario__gt=0).exclude(materia__asignaturamalla__ejeformativo_id__in=[4, 9, 11, 12]).count()
                        if cant_asignatura_malla == cant_materias_estudiante_total:
                            cant_materias_estudiante_nivel = MateriaAsignada.objects.filter(status=True, matricula=matricula, retiramateria=False, materia__asignaturamalla__nivelmalla=matricula.nivelmalla).exclude(materia__asignaturamalla__itinerario__gt=0).exclude(materia__asignaturamalla__ejeformativo_id__in=[4, 9, 11, 12]).count()
                            if cant_asignatura_malla == cant_materias_estudiante_nivel:
                                if matricula.nivelmalla.id == 1:
                                    if carrera_admision := list(MatriculacionPrimerNivelCarrera.objects.values_list('carreraadmision_id', flat=True).filter(status=True, carrerapregrado_id=carrera, configuracion_id=id_configuracion_admision_pregrado).distinct()):
                                        if materia_asignada_admision := MateriaAsignada.objects.values_list('notafinal', flat=True).filter(status=True, matricula__inscripcion__persona__cedula=matricula.inscripcion.persona.cedula, matricula__inscripcion__carrera__id=carrera_admision[0], estado=1):
                                            notarecord = round(sum(materia_asignada_admision) / materia_asignada_admision.count(), 2)
                                else:
                                    notarecord = matricula.inscripcion.promedio_record()

                                try:
                                    eInscripcion, ePeriodo = matricula.inscripcion, periodo
                                    eMatriculaAnterior = Matricula.objects.filter(status=True, nivel__periodo=ePeriodo, inscripcion=eInscripcion).first()
                                    regular, nivel_real = True, 1
                                    if eMatriculaAnterior:
                                        eInscripcion_mat = eMatriculaAnterior.inscripcion
                                        nivel_real = eMatriculaAnterior.nivelmalla.orden
                                    else:
                                        eInscripcion_mat = eInscripcion
                                        nivel_real = eInscripcion_mat.mi_nivel().nivel.orden
                                    EJE_FORMATIVO_VINCULACION, EJE_FORMATIVO_PRACTICAS = 11, 9
                                    EXCLUDE_EJE_FORMATIVO = [EJE_FORMATIVO_PRACTICAS, EJE_FORMATIVO_VINCULACION]
                                    eInscripcionMalla_mat = eInscripcion_mat.malla_inscripcion()
                                    eMalla_mat = eInscripcionMalla_mat.malla
                                    itinerario = []
                                    itinerario.append(0)
                                    if eInscripcionMalla_mat.malla.modalidad.id == 3:
                                        eAsignaturasMalla = eMalla_mat.asignaturamalla_set.select_related().filter(status=True).exclude(Q(nivelmalla_id=0) | Q(opcional=True)).exclude(itinerario__gt=0).order_by('nivelmalla', 'ejeformativo')
                                    else:
                                        eAsignaturasMalla = eMalla_mat.asignaturamalla_set.select_related().filter(status=True).exclude(Q(nivelmalla_id=0) | Q(opcional=True) | Q(ejeformativo_id__in=EXCLUDE_EJE_FORMATIVO)).exclude(itinerario__gt=0).order_by('nivelmalla', 'ejeformativo')
                                    eModulosMalla = ModuloMalla.objects.filter(malla=eInscripcionMalla_mat.malla)
                                    eNivelMallaAprobadoActual = eInscripcion_mat.recordacademico_set.filter(asignaturamalla__in=eAsignaturasMalla, status=True, aprobada=True).values_list('asignaturamalla__nivelmalla__orden', flat=True).order_by('-asignaturamalla__nivelmalla__orden').first()
                                    if eNivelMallaAprobadoActual:
                                        eAsignaturasMallaquedebeaprobar = eAsignaturasMalla.filter(nivelmalla__orden__lte=eNivelMallaAprobadoActual).exclude(Q(asignatura__id__in=eModulosMalla.values_list('asignatura__id', flat=True)) | Q(asignatura__in=AsignaturaMalla.objects.values_list("asignatura__id", flat=True).filter(malla=eMalla_mat)))
                                        eAsignaturasRecord = eInscripcion_mat.recordacademico_set.filter(asignaturamalla__in=eAsignaturasMalla, status=True).exclude(asignaturamalla__itinerario__gt=0)
                                        if eAsignaturasRecord.values("id").filter(aprobada=False).exists():
                                            regular = False
                                        elif eAsignaturasRecord.filter(asignaturamalla__in=eAsignaturasMallaquedebeaprobar, aprobada=True, status=True).count() < eAsignaturasMallaquedebeaprobar.values("id").count():
                                            regular = False
                                        elif eAsignaturasRecord.filter(asignaturamalla__in=eAsignaturasMalla.filter(nivelmalla__orden__gte=eNivelMallaAprobadoActual + 1).exclude(Q(asignatura__id__in=eModulosMalla.values_list('asignatura__id', flat=True)) | Q(asignatura__in=AsignaturaMalla.objects.values_list("asignatura__id", flat=True).filter(malla=eMalla_mat))), aprobada=True, status=True).exists():
                                            regular = False
                                        if eInscripcionMalla_mat.malla.modalidad.id == 3:
                                            eAsignaturasMallaHastaNivelActualTodas = eMalla_mat.asignaturamalla_set.select_related().filter(status=True, itinerario__in=itinerario, nivelmalla__orden__lte=eNivelMallaAprobadoActual).exclude(Q(nivelmalla_id=NIVEL_MALLA_CERO) | Q(opcional=True)).exclude(itinerario__gt=0)
                                        else:
                                            eAsignaturasMallaHastaNivelActualTodas = eMalla_mat.asignaturamalla_set.select_related().filter(status=True, itinerario__in=itinerario, nivelmalla__orden__lte=eNivelMallaAprobadoActual).exclude(Q(nivelmalla_id=NIVEL_MALLA_CERO) | Q(opcional=True) | Q(ejeformativo_id__in=EXCLUDE_EJE_FORMATIVO)).exclude(itinerario__gt=0)
                                        asignaturas_sin_aprobar = eAsignaturasMallaHastaNivelActualTodas.exclude(id__in=eAsignaturasRecord.values_list("asignaturamalla_id", flat=True))
                                        asignatura_esp = eAsignaturasMallaHastaNivelActualTodas.exclude(id__in=eAsignaturasRecord.values_list("asignaturamalla_id", flat=True)).filter(id__in=list(map(int, variable_valor('ASIGNATURAS_ESPECIALES'))))
                                        if asignaturas_sin_aprobar.exists():
                                            regular = False
                                        elif eAsignaturasRecord.values("id").filter(aprobada=False).exists():
                                            regular = False
                                        nivel_real = nivel_real + 1 if regular else nivel_real
                                        if regular and eMatriculaAnterior:
                                            niv_record = eNivelMallaAprobadoActual
                                            if niv_record:
                                                nivel_real = niv_record + 1
                                    tienerecordhomologadas = eInscripcion.recordacademico_set.values('id').filter(homologada=True, status=True).exists()
                                    if not eMatriculaAnterior and not tienerecordhomologadas:
                                        regular, nivel_real = True, 1
                                    es_estudiante_regular = regular
                                except Exception as ex:
                                    es_estudiante_regular = False

                                not es_estudiante_regular and print(f"{i}.- {matricula.inscripcion.persona} - {es_estudiante_regular} - {nivel[0]}".upper())

                                if es_estudiante_regular:
                                    materiasasignadas = MateriaAsignada.objects.filter(status=True, matricula=matricula, retiramateria=False)
                                    bandera = False
                                    if materiasasignadas:
                                        for materiaasignada in materiasasignadas:
                                            if materiaasignada.cantidad_matriculas() > 1:
                                                bandera = True
                                                break
                                        if not bandera:
                                            paralelos_estudiante = list(materiasasignadas.values_list('materia__paralelo', flat=True))
                                            contador = Counter(paralelos_estudiante)
                                            paralelo_perteneciente = contador.most_common(1)[0][0]
                                            inscripciones.append(matricula.nivelmalla.nombre)
                                            sexo = 'MUJER' if matricula.inscripcion.persona.es_mujer() else 'HOMBRE'
                                            if not EstudiantesCandidatosaPresidentesdeCurso.objects.filter(periodo_id=periodo.id, status=True, matricula=matricula.id, inscripcion=matricula.inscripcion.pk, carrera_id=carrera.pk).exists():
                                                presidentes = EstudiantesCandidatosaPresidentesdeCurso(status=True,
                                                                                                       periodo_id=periodo.id,
                                                                                                       periodo_nombre=periodo.nombre,
                                                                                                       carrera_id=carrera.pk,
                                                                                                       inscripcion=matricula.inscripcion.pk,
                                                                                                       matricula=matricula.pk,
                                                                                                       cedula=matricula.inscripcion.persona.cedula,
                                                                                                       nombres=matricula.inscripcion.persona.nombre_completo_inverso(),
                                                                                                       correo=matricula.inscripcion.persona.emailinst,
                                                                                                       telefono=matricula.inscripcion.persona.telefono,
                                                                                                       sexo=sexo,
                                                                                                       carrera=matricula.inscripcion.carrera.nombre,
                                                                                                       orden=int(nivel[0]),
                                                                                                       nivel=str(nivel[1]),
                                                                                                       paralelo=f"{paralelo_perteneciente}",
                                                                                                       promedio_final=str(notarecord))
                                                presidentes.save()
                                                contador_registro += 1
            except Exception as ex:
                transaction.set_rollback()
                raise
            noti = Notificacion(titulo=f'{carrera}', cuerpo=f'{carrera} ejecutada correctamente...', destinatario_id=37121, url="", prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1), tipo=2, en_proceso=False, error=True)
            noti.save()
    except Exception as ex:
        import sys
        noti = Notificacion(titulo='Error', cuerpo='Ha ocurrido un error {} - Error en la linea {}'.format(ex, sys.exc_info()[-1].tb_lineno), destinatario_id=37121, url="", prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1), tipo=2, en_proceso=False, error=True)
        noti.save()
        transaction.set_rollback(True)
        raise


def datauser():
    try:
        from datetime import time
        data = []
        carreras = Carrera.objects.filter(id__in=[137, 111])
        materias = Materia.objects.filter(nivel__periodo=336, asignaturamalla__malla__carrera__in=carreras, status=True).values('id')
        materiafiltrada = Clase.objects.values_list('materia', flat=True).filter(dia=2, materia__in=materias, turno__comienza__gte=time(13,0), turno__termina__lte=time(14,0), status=True).order_by('materia').distinct('materia')
        for materia in Materia.objects.filter(id__in=materiafiltrada):
            data.append([materia, materia.materiaasignada_set.filter(matricula__inscripcion__persona__sexo=1, status=True)])
        return data
    except Exception as ex:
        pass

# from sga.models import Periodo, Carrera, Materia
# from inno.models import EstudiantesCandidatosaPresidentesdeCurso
# def carga_de_candidatos_a_presidentes_actualizado_v3(periodo, carrera, niveles=None):
#     import sys
#     from settings import DEBUG
#     from collections import Counter
#     from django.db import transaction
#     from datetime import datetime, timedelta
#     from inno.models import EstudiantesCandidatosaPresidentesdeCurso
#     from sga.models import MatriculacionPrimerNivelCarrera, AsignaturaMalla, Matricula, MateriaAsignada, Notificacion
#     try:
#         notarecord = 0.0
#         inscripciones = []
#         bandera, matriculas = False, None
#         carrera_admision, contador_registro = 0, 0
#         with transaction.atomic():
#             try:
#                 # ELIMINAR LISTADO ANTERIOR PARA REALIZAR ALGUNA MODIFICACION
#                 listadonomigrados = '<br>'
#                 filtrobaseasignaturamalla = Q(status=True, malla__carrera=carrera, malla__status=True)
#                 if niveles:
#                     filtrobaseasignaturamalla &= Q(nivelmalla__orden__in=niveles)
#                 niveles_carrera = AsignaturaMalla.objects.values_list('nivelmalla__orden', 'nivelmalla__nombre').filter(filtrobaseasignaturamalla).order_by('nivelmalla__orden').distinct()
#                 filtrobasemateriaasignada = Q(materia__asignaturamalla__itinerario=0, retiromanual=False, retiramateria=False, matricula__retiradomatricula=False, matricula__status=True, materia__status=True, materia__asignaturamalla__status=True, materia__asignaturamalla__asignatura__modulo=False, status=True)
#                 for nivel in niveles_carrera:
#                     if nivel and not f"{nivel[0]}".isdigit():
#                         continue
#                     matriculas = Matricula.objects.filter(status=True, nivel__periodo=periodo, inscripcion__carrera_id=carrera, nivelmalla__orden=nivel[0]).order_by('nivelmalla__orden')
#                     listadonomigrados = '<br>'
#                     for i, matricula in enumerate(matriculas, start=1):
#                         try:
#                             mi_malla = matricula.inscripcion.mi_malla()
#                             materiasasignadas = matricula.materiaasignada_set.filter(filtrobasemateriaasignada & Q(materia__asignaturamalla__nivelmalla=matricula.nivelmalla)).exclude(materia__asignaturamalla__ejeformativo_id__in=[4, 9, 11, 12]).exclude(materia__asignaturamalla__malla=353)
#                             asignaturasmallae = mi_malla.asignaturamalla_set.filter(itinerario=0, asignatura__modulo=False, status=True, malla__carrera=matricula.inscripcion.carrera, nivelmalla=matricula.nivelmalla, vigente=True, opcional=False).exclude(ejeformativo_id__in=[4, 9, 11, 12])
#                             cant_asignatura_malla, cant_asignatura_estud = asignaturasmallae.count(), materiasasignadas.count()
#                             if cant_asignatura_malla == cant_asignatura_estud or matricula.tipomatriculalumno() == 'REGULAR':
#                                 tienesegundamatricula = False
#                                 # for materiaasignada in materiasasignadas:
#                                 #     if materiaasignada.cantidad_matriculas() > 1:
#                                 #         tienesegundamatricula = True
#                                 #         break
#                                 if materiasasignadas and not tienesegundamatricula:
#                                     paralelos_estudiante = list(materiasasignadas.values_list('materia__paralelo', flat=True))
#                                     contador = Counter(paralelos_estudiante)
#                                     paralelo_perteneciente = contador.most_common(1)[0][0]
#                                     inscripciones.append(matricula.nivelmalla.nombre)
#                                     sexo = 'MUJER' if matricula.inscripcion.persona.es_mujer() else 'HOMBRE'
#                                     if not EstudiantesCandidatosaPresidentesdeCurso.objects.filter(periodo_id=periodo.id, cedula=matricula.inscripcion.persona.cedula, nivel=str(nivel[1]), carrera_id=carrera.pk, status=True).exists():
#                                         if matricula and matricula.nivelmalla and matricula.nivelmalla.id == 1:
#                                             if id_configuracion_admision_pregrado := MatriculacionPrimerNivelCarrera.objects.values_list('configuracion_id', flat=True).filter(status=True, ejecutoaccion=True).order_by('configuracion_id').last():
#                                                 if carrera_admision := list(MatriculacionPrimerNivelCarrera.objects.values_list('carreraadmision_id', flat=True).filter(status=True, carrerapregrado_id=carrera, configuracion_id=id_configuracion_admision_pregrado).distinct()):
#                                                     if materia_asignada_admision := MateriaAsignada.objects.values_list('notafinal', flat=True).filter(status=True, matricula__inscripcion__persona__cedula=matricula.inscripcion.persona.cedula, matricula__inscripcion__carrera__id=carrera_admision[0], estado=1):
#                                                         notarecord = round(sum(materia_asignada_admision) / materia_asignada_admision.count(), 2)
#                                         else:
#                                             notarecord = matricula.inscripcion.promedio_record()
#                                         presidentes = EstudiantesCandidatosaPresidentesdeCurso(status=True,
#                                                                                                periodo_id=periodo.id,
#                                                                                                periodo_nombre=periodo.nombre,
#                                                                                                carrera_id=carrera.pk,
#                                                                                                inscripcion=matricula.inscripcion.pk,
#                                                                                                matricula=matricula.pk,
#                                                                                                cedula=matricula.inscripcion.persona.cedula,
#                                                                                                nombres=matricula.inscripcion.persona.nombre_completo_inverso(),
#                                                                                                correo=matricula.inscripcion.persona.emailinst,
#                                                                                                telefono=matricula.inscripcion.persona.telefono,
#                                                                                                sexo=sexo,
#                                                                                                carrera=matricula.inscripcion.carrera.nombre,
#                                                                                                orden=int(nivel[0]),
#                                                                                                nivel=str(nivel[1]),
#                                                                                                paralelo=f"{paralelo_perteneciente}",
#                                                                                                promedio_final=str(notarecord))
#                                         presidentes.save()
#                                         contador_registro += 1
#                                         DEBUG and print(f"{presidentes.nivel}.- {matricula.inscripcion.persona.nombre_completo_inverso()} - {carrera.nombre}")
#                         except Exception as ex:
#                             listadonomigrados += f'{matricula.inscripcion.persona} - {ex} <br>'
#                 noti = Notificacion(titulo=f'Proceso finalizado...', cuerpo=f'{listadonomigrados} - {carrera}', destinatario_id=37121, url="", prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1), tipo=2, en_proceso=False, error=True)
#                 noti.save()
#             except Exception as ex:
#                 transaction.set_rollback(True)
#                 raise NameError(f"{ex} - L: {sys.exc_info()[-1].tb_lineno}")
#     except Exception as ex:
#         noti = Notificacion(titulo=f'Error en {carrera}'.upper(), cuerpo=f'Ha ocurrido un error {ex} - error en la linea {sys.exc_info()[-1].tb_lineno}', destinatario_id=37121, url="", prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1), tipo=2, en_proceso=False, error=True)
#         noti.save()
#
# periodo = Periodo.objects.get(id=336)
# __carrera = Materia.objects.values_list('asignaturamalla__malla__carrera', flat=True).filter(status=True, nivel__periodo=periodo).distinct()
# for carrera in Carrera.objects.filter(id__in=[133], coordinacion__in=[1,2,3,4,5], status=True, malla__vigente=True):
#     # EstudiantesCandidatosaPresidentesdeCurso.objects.filter(carrera_id=carrera.pk, periodo_id=periodo.pk).delete()
#     carga_de_candidatos_a_presidentes_actualizado_v3(periodo, carrera, [2])
#     #funcionlocal.get("carga_de_candidatos_a_presidentes_actualizado_v3")(periodo, carrera)
#
# # carrerasconitinerario = Materia.objects.values_list('asignaturamalla__malla__carrera__id','asignaturamalla__nivelmalla__orden').filter(status=True, nivel__periodo=periodo).exclude(asignaturamalla__itinerario=0).distinct()
# # carreras = list(set([c[0] for c in carrerasconitinerario]))
# # for c in carreras:
# #     niveles = [x[1] for x in carrerasconitinerario if x[0] == c]
# #     carrera = Carrera.objects.get(id=c)
# #     #carga_de_candidatos_a_presidentes_actualizado_v3(periodo, carrera, niveles)
# #     funcionlocal.get("carga_de_candidatos_a_presidentes_actualizado_v3")(periodo, carrera, niveles)
