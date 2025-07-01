# -*- coding: latin-1 -*-
import random
import sys
from datetime import datetime, timedelta

import xlrd
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models.query_utils import Q
from django.http import HttpResponseRedirect, JsonResponse, HttpResponse
from django.shortcuts import render
from django.template.loader import get_template
from googletrans import Translator
from xlwt import *

from decorators import secure_module, last_access
from sagest.models import DistributivoPersona
from sga.commonviews import adduserdata, traerNotificaciones
from sga.excelbackground import reporte_encuesta_grupo_estudiante_background
from sga.forms import EncuestaGrupoEstudiantesForm, \
    PreguntaEncuestaGrupoEstudiantesForm, RangoPreguntaEncuestaGrupoEstudiantesForm, \
    OpcionCuadriculaEncuestaGrupoEstudiantesForm, OpcionMultipleEncuestaGrupoEstudiantesForm, \
    DuplicarContenidoPreguntaForm, DuplicarContenidoEncuestaForm
from sga.funciones import log, MiPaginador
from sga.models import EncuestaGrupoEstudiantes, PreguntaEncuestaGrupoEstudiantes, \
    RangoPreguntaEncuestaGrupoEstudiantes, OpcionCuadriculaEncuestaGrupoEstudiantes, \
    OpcionMultipleEncuestaGrupoEstudiantes, ProfesorDistributivoHoras, Inscripcion, Administrativo, Profesor, \
    InscripcionEncuestaGrupoEstudiantes, Notificacion, Persona
from inno.models import EncuestaGrupoEstudianteSeguimientoSilabo


@login_required(redirect_field_name='ret', login_url='/loginsga')
@secure_module
@last_access
@transaction.atomic()
def view(request):
    data = {}
    adduserdata(request, data)
    persona = request.session['persona']
    if request.method == 'POST':
        action = request.POST['action']

        if action == 'add':
            try:
                f = EncuestaGrupoEstudiantesForm(request.POST)
                if f.is_valid():
                    if EncuestaGrupoEstudiantes.objects.filter(status=True, descripcion=f.cleaned_data['descripcion'], tipoperfil=f.cleaned_data['tipoperfil']).exists():
                        return JsonResponse({"result": "bad", "mensaje": u"La encuesta ya existe."})
                    encuesta = EncuestaGrupoEstudiantes(descripcion=f.cleaned_data['descripcion'],
                                                        tipoperfil=f.cleaned_data['tipoperfil'],
                                                        activo=f.cleaned_data['activo'],
                                                        leyenda=f.cleaned_data['leyenda'],
                                                        obligatoria=f.cleaned_data['obligatoria'],
                                                        )
                    encuesta.save(request)
                    #APARTADO PARA CATEGORIZAR LA ENCUESTA
                    if f.cleaned_data['categoria']:
                        # encuesta.activo = False
                        encuesta_seguimiento = EncuestaGrupoEstudianteSeguimientoSilabo(
                            encuestagrupoestudiantes_id = encuesta.pk,
                            categoria = f.cleaned_data['categoria'])
                        encuesta_seguimiento.save(request)
                        # encuesta.save(request)
                    log(u'Adiciono encuesta grupo estudiante: %s' % encuesta, request, "add")
                    return JsonResponse({"result": "ok", "id": encuesta.id})
                else:
                     raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                translator = Translator()
                return JsonResponse({"result": "bad", "ex": translator.translate(ex.__str__(),'es').text, "mensaje": u"Error al guardar los datos."})

        elif action == 'eliminar_inscripcion':
            try:
                res_json = []
                with transaction.atomic():
                    instancia = InscripcionEncuestaGrupoEstudiantes.objects.get(pk=int(request.POST['id']))
                    instancia.delete()
                    log(u'Elimino Inscripciòn encuesta grupo: %s' % instancia, request, "delete")
                    res_json = {"error": False}
            except Exception as ex:
                res_json = {'error': True, "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)

        elif action == 'addpregunta':
            try:
                f = PreguntaEncuestaGrupoEstudiantesForm(request.POST)
                if f.is_valid():
                    if int(f.cleaned_data['tipo']) == 1:
                        preguntaencuestagrupoestudiantes = PreguntaEncuestaGrupoEstudiantes(encuesta_id=int(request.POST['idencuesta']),
                                                                    tipo=f.cleaned_data['tipo'],
                                                                    descripcion=f.cleaned_data['descripcion'],
                                                                    orden=f.cleaned_data['orden'],
                                                                    observacionporno=f.cleaned_data['observacionporno'],
                                                                    obligatoria=f.cleaned_data['obligatoria'])
                    else:
                        preguntaencuestagrupoestudiantes = PreguntaEncuestaGrupoEstudiantes(encuesta_id=int(request.POST['idencuesta']),
                                                                    tipo=f.cleaned_data['tipo'],
                                                                    descripcion=f.cleaned_data['descripcion'],
                                                                    orden=f.cleaned_data['orden'],
                                                                    obligatoria=f.cleaned_data['obligatoria'])
                    preguntaencuestagrupoestudiantes.save(request)
                    log(u'Adiciono Pregunta encuesta grupo estudiante: %s' % preguntaencuestagrupoestudiantes, request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                     raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                translator = Translator()
                return JsonResponse({"result": "bad", "ex": translator.translate(ex.__str__(),'es').text, "mensaje": u"Error al guardar los datos."})

        elif action == 'addrangopregunta':
            try:
                f = RangoPreguntaEncuestaGrupoEstudiantesForm(request.POST)
                if f.is_valid():
                    rangopreguntaencuestagrupoestudiantes = RangoPreguntaEncuestaGrupoEstudiantes(pregunta_id=int(request.POST['idpregunta']),
                                                                                             descripcion=f.cleaned_data['descripcion'],
                                                                                             orden=f.cleaned_data['orden'],
                                                                                             valor=f.cleaned_data['valor'])
                    rangopreguntaencuestagrupoestudiantes.save(request)
                    log(u'Adiciono Rango Pregunta encuesta grupo estudiante: %s' % rangopreguntaencuestagrupoestudiantes, request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                     raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                translator = Translator()
                return JsonResponse({"result": "bad", "ex": translator.translate(ex.__str__(),'es').text, "mensaje": u"Error al guardar los datos."})

        elif action == 'addopcioncuadricula':
            try:
                f = OpcionCuadriculaEncuestaGrupoEstudiantesForm(request.POST)
                if f.is_valid():
                    opcioncuadricula = OpcionCuadriculaEncuestaGrupoEstudiantes(pregunta_id=int(request.POST['idpregunta']),
                                                                                descripcion=f.cleaned_data['descripcion'],
                                                                                orden=f.cleaned_data['orden'],
                                                                                valor=f.cleaned_data['valor'],
                                                                                tipoopcion=int(request.POST['tipo']),
                                                                                opcotros=f.cleaned_data['otros'],
                                                                                oparchivo=f.cleaned_data['archivo'],
                                                                                secuenciapregunta=f.cleaned_data['secuenciapregunta']
                                                                                )

                    opcioncuadricula.save(request)
                    log(u'Adiciono opción cuadrícula de encuesta grupo estudiante: %s' % opcioncuadricula, request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                     raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                translator = Translator()
                return JsonResponse({"result": "bad", "ex": translator.translate(ex.__str__(),'es').text, "mensaje": u"Error al guardar los datos."})

        elif action == 'edit':
            try:
                encuesta = EncuestaGrupoEstudiantes.objects.get(pk=request.POST['id'])
                f = EncuestaGrupoEstudiantesForm(request.POST)
                if f.is_valid():
                    if EncuestaGrupoEstudiantes.objects.filter(status=True, descripcion=f.cleaned_data['descripcion'], tipoperfil=f.cleaned_data['tipoperfil']).exclude(pk=encuesta.id).exists():
                        return JsonResponse({"result": "bad", "mensaje": u"La Encuesta ya existe."})
                    encuesta.descripcion = f.cleaned_data['descripcion']
                    encuesta.leyenda = f.cleaned_data['leyenda']
                    encuesta.tipoperfil = f.cleaned_data['tipoperfil']
                    encuesta.activo = f.cleaned_data['activo']
                    encuesta.obligatoria = f.cleaned_data['obligatoria']
                    encuesta.save(request)
                    if not EncuestaGrupoEstudianteSeguimientoSilabo.objects.filter(
                            encuestagrupoestudiantes_id=encuesta.id).exists():
                        if f.cleaned_data['categoria']:
                            # encuesta.activo = False
                            encuesta_seguimiento = EncuestaGrupoEstudianteSeguimientoSilabo(
                                encuestagrupoestudiantes_id=encuesta.pk,
                                categoria=f.cleaned_data['categoria'])
                            encuesta_seguimiento.save(request)
                            # encuesta.save(request)
                    else:
                        encuesta_seguimiento = EncuestaGrupoEstudianteSeguimientoSilabo.objects.get(
                            encuestagrupoestudiantes_id=request.POST['id'])
                        if f.cleaned_data['categoria']:
                            # encuesta.activo = False
                            encuesta_seguimiento.categoria = f.cleaned_data['categoria']
                            # encuesta.save(request)
                        else:
                            encuesta_seguimiento.categoria = False
                        encuesta_seguimiento.save(request)
                    log(u'Modifico encuesta grupo estudiante: %s' % encuesta, request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                     raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                translator = Translator()
                return JsonResponse({"result": "bad", "ex": translator.translate(ex.__str__(),'es').text, "mensaje": u"Error al guardar los datos."})

        elif action == 'addopcionmultiple':
            try:
                f = OpcionMultipleEncuestaGrupoEstudiantesForm(request.POST)
                if f.is_valid():
                    opcionmultiple = OpcionMultipleEncuestaGrupoEstudiantes(pregunta_id=int(request.POST['idpregunta']),
                                                                                descripcion=f.cleaned_data['descripcion'],
                                                                                orden=f.cleaned_data['orden'],
                                                                                valor=f.cleaned_data['valor'],
                                                                                opcotros = f.cleaned_data['otros'])

                    opcionmultiple.save(request)
                    log(u'Adiciono opción multiple de encuesta grupo estudiante: %s' % opcionmultiple, request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                     raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                translator = Translator()
                return JsonResponse({"result": "bad", "ex": translator.translate(ex.__str__(),'es').text, "mensaje": u"Error al guardar los datos."})

        elif action == 'editpregunta':
            try:
                pregunta = PreguntaEncuestaGrupoEstudiantes.objects.get(pk=request.POST['id'])
                f = PreguntaEncuestaGrupoEstudiantesForm(request.POST)
                if f.is_valid():
                    # if EncuestaGrupoEstudiantes.objects.filter(status=True, descripcion=f.cleaned_data['descripcion']).exclude(pk=encuesta.id).exists():
                    #     return JsonResponse({"result": "bad", "mensaje": u"La Encuesta ya existe."})
                    pregunta.descripcion = f.cleaned_data['descripcion']
                    pregunta.tipo = f.cleaned_data['tipo']
                    pregunta.orden = f.cleaned_data['orden']
                    pregunta.obligatoria = f.cleaned_data['obligatoria']
                    if int(f.cleaned_data['tipo']) == 1:
                        pregunta.observacionporno = f.cleaned_data['observacionporno']
                    pregunta.save(request)
                    log(u'Modifico pregunta encuesta grupo estudiante: %s' % pregunta, request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                     raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                translator = Translator()
                return JsonResponse({"result": "bad", "ex": translator.translate(ex.__str__(),'es').text, "mensaje": u"Error al guardar los datos."})

        elif action == 'editrangopregunta':
            try:
                rango = RangoPreguntaEncuestaGrupoEstudiantes.objects.get(pk=request.POST['id'])
                f = RangoPreguntaEncuestaGrupoEstudiantesForm(request.POST)
                if f.is_valid():
                    rango.descripcion = f.cleaned_data['descripcion']
                    rango.valor = f.cleaned_data['valor']
                    rango.orden = f.cleaned_data['orden']
                    rango.save(request)
                    log(u'Modifico Rango pregunta encuesta grupo estudiante: %s' % rango, request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                     raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                translator = Translator()
                return JsonResponse({"result": "bad", "ex": translator.translate(ex.__str__(),'es').text, "mensaje": u"Error al guardar los datos."})

        elif action == 'editopcioncuadricula':
            try:
                opcion = OpcionCuadriculaEncuestaGrupoEstudiantes.objects.get(pk=request.POST['id'])
                f = OpcionCuadriculaEncuestaGrupoEstudiantesForm(request.POST)
                if f.is_valid():
                    opcion.descripcion = f.cleaned_data['descripcion']
                    opcion.valor = f.cleaned_data['valor']
                    opcion.orden = f.cleaned_data['orden']
                    opcion.opcotros = f.cleaned_data['otros']
                    opcion.oparchivo = f.cleaned_data['archivo']
                    opcion.secuenciapregunta = f.cleaned_data['secuenciapregunta']
                    opcion.save(request)
                    log(u'Modifico opción de cuadrícula encuesta grupo estudiante: %s' % opcion, request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                     raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                translator = Translator()
                return JsonResponse({"result": "bad", "ex": translator.translate(ex.__str__(),'es').text, "mensaje": u"Error al guardar los datos."})

        elif action == 'editopcionmultiple':
            try:
                opcion = OpcionMultipleEncuestaGrupoEstudiantes.objects.get(pk=request.POST['id'])
                f = OpcionMultipleEncuestaGrupoEstudiantesForm(request.POST)
                if f.is_valid():
                    opcion.descripcion = f.cleaned_data['descripcion']
                    opcion.valor = f.cleaned_data['valor']
                    opcion.orden = f.cleaned_data['orden']
                    opcion.opcotros = f.cleaned_data['otros']
                    opcion.save(request)
                    log(u'Modifico opción de respuesta multiple encuesta grupo estudiante: %s' % opcion, request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                     raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                translator = Translator()
                return JsonResponse({"result": "bad", "ex": translator.translate(ex.__str__(),'es').text, "mensaje": u"Error al guardar los datos."})

        elif action == 'delete':
            try:
                encuesta = EncuestaGrupoEstudiantes.objects.get(pk=request.POST['id'])
                if encuesta.enuso():
                    return JsonResponse({"result": "bad", "mensaje": u"Encuesta en uso."})
                log(u'Elimino encuesta grupo estudiante: %s' % encuesta, request, "del")
                encuesta.delete()
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                translator = Translator()
                return JsonResponse({"result": "bad", "ex": translator.translate(ex.__str__(),'es').text, "mensaje": u"Error al eliminar los datos."})

        elif action == 'deletepregunta':
            try:
                pregunta = PreguntaEncuestaGrupoEstudiantes.objects.get(pk=request.POST['id'])
                if pregunta.enuso():
                    return JsonResponse({"result": "bad", "mensaje": u"Pregunta en uso."})
                log(u'Elimino Pregunta encuesta grupo estudiante: %s' % pregunta, request, "del")
                pregunta.delete()
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                translator = Translator()
                return JsonResponse({"result": "bad", "ex": translator.translate(ex.__str__(),'es').text, "mensaje": u"Error al eliminar los datos."})

        elif action == 'deleterangopregunta':
            try:
                rango = RangoPreguntaEncuestaGrupoEstudiantes.objects.get(pk=request.POST['id'])
                if rango.enuso():
                    return JsonResponse({"result": "bad", "mensaje": u"Rango en uso."})
                log(u'Elimino rango pregunta encuesta grupo estudiante: %s' % rango, request, "del")
                rango.delete()
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                translator = Translator()
                return JsonResponse({"result": "bad", "ex": translator.translate(ex.__str__(),'es').text, "mensaje": u"Error al eliminar los datos."})

        elif action == 'deleteopcioncuadricula':
            try:
                opcion = OpcionCuadriculaEncuestaGrupoEstudiantes.objects.get(pk=request.POST['id'])
                # if opcion.enuso():
                #     return JsonResponse({"result": "bad", "mensaje": u"Opción de cuadrícula en uso."})

                opcion.delete()
                log(u'Elimino rango pregunta encuesta grupo estudiante: %s' % opcion, request, "del")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                translator = Translator()
                return JsonResponse({"result": "bad", "ex": translator.translate(ex.__str__(),'es').text, "mensaje": u"Error al eliminar los datos."})

        elif action == 'deleteopcionmultiple':
            try:
                opcion = OpcionMultipleEncuestaGrupoEstudiantes.objects.get(pk=request.POST['id'])
                if opcion.enuso():
                    return JsonResponse({"result": "bad", "mensaje": u"Opción multiple en uso."})
                opcion.delete()
                log(u'Elimino opcion multiple grupo estudiante: %s' % opcion, request, "del")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                translator = Translator()
                return JsonResponse({"result": "bad", "ex": translator.translate(ex.__str__(),'es').text, "mensaje": u"Error al eliminar los datos."})

        elif action == 'importarPoblacion':
            try:
                import openpyxl
                if 'archivo_excel' in request.FILES:
                    archivo_excel = request.FILES['archivo_excel']
                    if archivo_excel:
                        id_encuesta = request.POST['id']
                        encuesta = EncuestaGrupoEstudiantes.objects.get(pk=id_encuesta, status=True)
                        workbook = openpyxl.load_workbook(archivo_excel)
                        sheet = workbook.active
                        m_row = sheet.max_row-1
                        for i in range(1, m_row + 1):
                            celda = sheet.cell(row=i+1, column=1)
                            if ((sheet.cell(row=1, column=1)).value == 'id'):
                                if not celda.value == None:
                                    if encuesta.tipoperfil == 1:  # ESTUDIANTE
                                        if Inscripcion.objects.filter(pk=int(celda.value)).exists():
                                            inscripcion =InscripcionEncuestaGrupoEstudiantes.objects.filter(inscripcion_id = int(celda.value), encuesta=encuesta)
                                            if inscripcion.exists():
                                                for dato in inscripcion:
                                                    if not dato.status:
                                                        dato.status =True
                                                        dato.save()
                                                        log(u'Se actualizo la inscripcion encuesta grupo: %s' % dato, request, "edit")
                                            else:
                                                registro = InscripcionEncuestaGrupoEstudiantes(
                                                    encuesta=encuesta,
                                                    inscripcion_id=int(celda.value)
                                                )
                                                registro.save()
                                                log(u'Se guardo la inscripcion encuesta grupo: %s' % registro, request, "add")
                                    if encuesta.tipoperfil == 2:  # DOCENTE
                                        if Profesor.objects.filter(pk=int(celda.value)).exists():
                                            inscripcion = InscripcionEncuestaGrupoEstudiantes.objects.filter(profesor_id=int(celda.value), encuesta=encuesta)
                                            if inscripcion.exists():
                                                for dato in inscripcion:
                                                    if not dato.status:
                                                        dato.status =True
                                                        dato.save()
                                                        log(u'Se actualizo la inscripcion encuesta grupo: %s' % dato, request, "edit")
                                            else:
                                                registro = InscripcionEncuestaGrupoEstudiantes(
                                                    encuesta=encuesta,
                                                    profesor_id=int(celda.value)
                                                )
                                                registro.save()
                                                log(u'Se guardo la inscripcion encuesta grupo: %s' % registro, request, "add")
                                    if encuesta.tipoperfil == 3:  # ADMINISTRATIVOS
                                        if Administrativo.objects.filter(pk=int(celda.value)).exists():
                                            inscripcion = InscripcionEncuestaGrupoEstudiantes.objects.filter(administrativo_id=int(celda.value), encuesta=encuesta)
                                            if inscripcion.exists():
                                                for dato in inscripcion:
                                                    if not dato.status:
                                                        dato.status = True
                                                        dato.save()
                                                        log(u'Se actualizo la inscripcion encuesta grupo: %s' % dato, request, "edit")
                                            else:
                                                registro = InscripcionEncuestaGrupoEstudiantes(
                                                    encuesta=encuesta,
                                                    administrativo_id=int(celda.value)
                                                )
                                                registro.save()
                                                log(u'Se guardo la inscripcion encuesta grupo: %s' % registro, request, "add")
                                    if encuesta.tipoperfil == 4:  # ADMINISTRATIVO - DOCENTES
                                        if Persona.objects.filter(pk=int(celda.value)).exists():
                                            inscripcion = InscripcionEncuestaGrupoEstudiantes.objects.filter(persona_id=int(celda.value), encuesta=encuesta)
                                            if inscripcion.exists():
                                                for dato in inscripcion:
                                                    if not dato.status:
                                                        dato.status = True
                                                        dato.save()
                                                        log(u'Se actualizo la inscripcion encuesta grupo: %s' % dato, request, "edit")
                                            else:
                                                registro = InscripcionEncuestaGrupoEstudiantes(
                                                    encuesta=encuesta,
                                                    persona_id=int(celda.value)
                                                )
                                                registro.save()
                                                log(u'Se guardo la inscripcion encuesta grupo: %s' % registro, request, "add")
                            else:
                                return JsonResponse({"result": True, "mensaje": u"El formato de las columnas del archivo no es el correcto."})

                    return JsonResponse({"result": False, "mensaje": u"Importaciòn realizada correctamente."}, safe=False)
                else:
                    return JsonResponse({"result": True, "mensaje": u"Seleccione una archivo"})
            except Exception as ex:
                transaction.set_rollback(True)
                translator = Translator()
                return JsonResponse({"result": "True", "ex": translator.translate(ex.__str__(), 'es').text, "mensaje": u"Error al guardar los datos."})

        elif action == 'saveAsignarPoblacion':
            try:
                id_encuesta = request.POST['ide']
                tipoperfil = request.POST['tipoperfil']
                id = int(request.POST['id'])
                encuesta = EncuestaGrupoEstudiantes.objects.get(pk=id_encuesta, status=True)
                if encuesta.tipoperfil == 1:  # ESTUDIANTE
                    if Inscripcion.objects.filter(pk=id).exists():
                        registros = InscripcionEncuestaGrupoEstudiantes.objects.filter(inscripcion_id=id, encuesta=encuesta)
                        if registros.exists():
                            for registro in registros:
                                if not registro.status:
                                    registro.status = True
                                    registro.save()
                                    log(u'Se actualizo la población de encuesta grupo: %s' % registro, request, "edit")
                        else:
                            registro = InscripcionEncuestaGrupoEstudiantes(encuesta=encuesta,
                                                                           inscripcion_id=id)
                            registro.save()
                            log(u'Se guardo la población de encuesta grupo: %s' % registro, request, "add")
                elif encuesta.tipoperfil == 2:  # DOCENTE
                    if Profesor.objects.filter(pk=id).exists():
                        registros = InscripcionEncuestaGrupoEstudiantes.objects.filter(profesor_id=id, encuesta=encuesta)
                        if registros.exists():
                            for registro in registros:
                                if not registro.status:
                                    registro.status =True
                                    registro.save()
                                    log(u'Se actualizo la población de encuesta grupo: %s' % registro, request, "edit")
                        else:
                            registro = InscripcionEncuestaGrupoEstudiantes(encuesta=encuesta,
                                                                           profesor_id=id)
                            registro.save()
                            log(u'Se guardo la población de encuesta grupo: %s' % registro, request, "add")
                elif encuesta.tipoperfil == 3:  # ADMINISTRATIVOS
                    if Administrativo.objects.filter(pk=id).exists():
                        registros = InscripcionEncuestaGrupoEstudiantes.objects.filter(administrativo_id=id, encuesta=encuesta)
                        if registros.exists():
                            for registro in registros:
                                if not registro.status:
                                    registro.status = True
                                    registro.save()
                                    log(u'Se actualizo la población de encuesta grupo: %s' % registro, request, "edit")
                        else:
                            registro = InscripcionEncuestaGrupoEstudiantes(encuesta=encuesta,
                                                                           administrativo_id=id)
                            registro.save()
                            log(u'Se guardo la población de encuesta grupo: %s' % registro, request, "add")
                return JsonResponse({"result": 'ok', "mensaje": u"Se guardo correctamente el registro."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos. %s" % ex.__str__()})

        elif action == 'duplicarconenidopregunta':
            try:
                ePreguntaEncuestaGrupoEstudiantes = PreguntaEncuestaGrupoEstudiantes.objects.get(pk=request.POST['id'])

                form = DuplicarContenidoPreguntaForm(request.POST)
                pregunta_clonar = PreguntaEncuestaGrupoEstudiantes.objects.get(pk=int(request.POST['pregunta']))
                if pregunta_clonar:
                    if ePreguntaEncuestaGrupoEstudiantes.tipo == 2:
                        if not RangoPreguntaEncuestaGrupoEstudiantes.objects.filter( pregunta_id=ePreguntaEncuestaGrupoEstudiantes.pk, status=True).exists():
                            eRangoPreguntaEncuestaGrupoEstudiantes = RangoPreguntaEncuestaGrupoEstudiantes.objects.filter(pregunta_id=pregunta_clonar.pk,status=True)
                            for d in eRangoPreguntaEncuestaGrupoEstudiantes:
                                eRangoPreguntaEncuestaGrupoEstudiantes = RangoPreguntaEncuestaGrupoEstudiantes(
                                    pregunta_id=ePreguntaEncuestaGrupoEstudiantes.pk,
                                    descripcion=d.descripcion,
                                    orden=d.orden,
                                    valor=d.valor
                                )
                                eRangoPreguntaEncuestaGrupoEstudiantes.save(request)
                                log(u'Adiciono Rango Pregunta encuesta grupo estudiante: %s' % eRangoPreguntaEncuestaGrupoEstudiantes,request, "add")

                    if ePreguntaEncuestaGrupoEstudiantes.tipo == 5:
                        if not OpcionCuadriculaEncuestaGrupoEstudiantes.objects.filter( pregunta_id=ePreguntaEncuestaGrupoEstudiantes.pk, status=True).exists():
                            eOpcionCuadriculaEncuestaGrupoEstudiantes = OpcionCuadriculaEncuestaGrupoEstudiantes.objects.filter(pregunta_id=pregunta_clonar.pk,status=True)
                            for d in eOpcionCuadriculaEncuestaGrupoEstudiantes:
                                opcioncuadricula = OpcionCuadriculaEncuestaGrupoEstudiantes(
                                    pregunta_id=ePreguntaEncuestaGrupoEstudiantes.pk,
                                    descripcion=d.descripcion,
                                    orden=d.orden,
                                    valor=d.valor,
                                    tipoopcion=d.tipoopcion,
                                    opcotros=d.opcotros,
                                    oparchivo=d.oparchivo,
                                    secuenciapregunta=d.secuenciapregunta
                                    )

                                opcioncuadricula.save(request)
                                log(u'Adiciono opción cuadrícula de encuesta grupo estudiante: %s' % opcioncuadricula,
                                    request, "add")

                    if ePreguntaEncuestaGrupoEstudiantes.tipo == 6:
                        if not OpcionMultipleEncuestaGrupoEstudiantes.objects.filter( pregunta_id=ePreguntaEncuestaGrupoEstudiantes.pk, status=True).exists():
                            eOpcionMultipleEncuestaGrupoEstudiantes = OpcionMultipleEncuestaGrupoEstudiantes.objects.filter( pregunta_id=pregunta_clonar.pk, status=True)
                            for d in eOpcionMultipleEncuestaGrupoEstudiantes:
                                opcionmultiple = OpcionMultipleEncuestaGrupoEstudiantes(
                                    pregunta_id=ePreguntaEncuestaGrupoEstudiantes.pk,
                                    descripcion=d.descripcion,
                                    orden=d.orden,
                                    valor=d.valor,
                                    opcotros=d.opcotros)
                                opcionmultiple.save(request)
                                log(u'Adiciono opción multiple de encuesta grupo estudiante: %s' % opcionmultiple, request,
                                    "add")

                return JsonResponse({"result": False, "mensaje": u"Se guardo correctamente el registro."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos. %s" % ex.__str__()})

        elif action == 'duplicarcontenidoencuesta':
            try:

                eEncuestaGrupoEstudiantes = EncuestaGrupoEstudiantes.objects.get(pk=int(request.POST['id']))
                encuesta_a_clonar = EncuestaGrupoEstudiantes.objects.get(pk=int(request.POST['encuesta']))
                for ePreguntaEncuestaGrupoEstudiantes in encuesta_a_clonar.get_preguntas():
                    pregunta_clonar = ePreguntaEncuestaGrupoEstudiantes
                    nuevapregunta =PreguntaEncuestaGrupoEstudiantes(
                        encuesta = eEncuestaGrupoEstudiantes,
                        tipo = pregunta_clonar.tipo,
                        descripcion  = pregunta_clonar.descripcion,
                        observacionporno  = pregunta_clonar.observacionporno,
                        orden = pregunta_clonar.orden,
                        obligatoria =pregunta_clonar.obligatoria
                    )
                    nuevapregunta.save(request)
                    if nuevapregunta:
                        if nuevapregunta.tipo == 2:
                            if not RangoPreguntaEncuestaGrupoEstudiantes.objects.filter(pregunta_id=nuevapregunta.pk, status=True).exists():
                                eRangoPreguntaEncuestaGrupoEstudiantes = RangoPreguntaEncuestaGrupoEstudiantes.objects.filter(pregunta_id=pregunta_clonar.pk,status=True)
                                for d in eRangoPreguntaEncuestaGrupoEstudiantes:
                                    eRangoPreguntaEncuestaGrupoEstudiantes = RangoPreguntaEncuestaGrupoEstudiantes(
                                        pregunta_id=nuevapregunta.pk,
                                        descripcion=d.descripcion,
                                        orden=d.orden,
                                        valor=d.valor
                                    )
                                    eRangoPreguntaEncuestaGrupoEstudiantes.save(request)
                                    log(u'Adiciono Rango Pregunta encuesta grupo estudiante: %s' % eRangoPreguntaEncuestaGrupoEstudiantes,request, "add")

                        if nuevapregunta.tipo == 5:
                            if not OpcionCuadriculaEncuestaGrupoEstudiantes.objects.filter( pregunta_id=nuevapregunta.pk, status=True).exists():
                                eOpcionCuadriculaEncuestaGrupoEstudiantes = OpcionCuadriculaEncuestaGrupoEstudiantes.objects.filter(pregunta_id=pregunta_clonar.pk,status=True)
                                for d in eOpcionCuadriculaEncuestaGrupoEstudiantes:
                                    opcioncuadricula = OpcionCuadriculaEncuestaGrupoEstudiantes(
                                        pregunta_id=nuevapregunta.pk,
                                        descripcion=d.descripcion,
                                        orden=d.orden,
                                        valor=d.valor,
                                        tipoopcion=d.tipoopcion,
                                        opcotros=d.opcotros,
                                        oparchivo=d.oparchivo,
                                        secuenciapregunta=d.secuenciapregunta
                                        )

                                    opcioncuadricula.save(request)
                                    log(u'Adiciono opción cuadrícula de encuesta grupo estudiante: %s' % opcioncuadricula,
                                        request, "add")

                        if nuevapregunta.tipo == 6:
                            if not OpcionMultipleEncuestaGrupoEstudiantes.objects.filter( pregunta_id=nuevapregunta.pk, status=True).exists():
                                eOpcionMultipleEncuestaGrupoEstudiantes = OpcionMultipleEncuestaGrupoEstudiantes.objects.filter( pregunta_id=pregunta_clonar.pk, status=True)
                                for d in eOpcionMultipleEncuestaGrupoEstudiantes:
                                    opcionmultiple = OpcionMultipleEncuestaGrupoEstudiantes(
                                        pregunta_id=nuevapregunta.pk,
                                        descripcion=d.descripcion,
                                        orden=d.orden,
                                        valor=d.valor,
                                        opcotros=d.opcotros)
                                    opcionmultiple.save(request)
                                    log(u'Adiciono opción multiple de encuesta grupo estudiante: %s' % opcionmultiple, request,
                                        "add")

                return JsonResponse({"result":False, "mensaje": u"Se guardo correctamente el registro."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos. %s" % ex.__str__()})

        return JsonResponse({"result": "bad", "mensaje": u"Solicitud Incorrecta."})
    else:
        data['title'] = u'Listado de Encuesta'
        if 'action' in request.GET:
            action = request.GET['action']

            if action == 'add':
                try:
                    data['title'] = u'Adicionar encuesta'
                    data['form'] = EncuestaGrupoEstudiantesForm(initial={})
                    return render(request, "adm_encuestagrupoestudiantes/add.html", data)
                except Exception as ex:
                    pass

            if action == 'addpregunta':
                try:
                    data['title'] = u'Adicionar Pregunta encuesta'
                    data['form'] = PreguntaEncuestaGrupoEstudiantesForm(initial={})
                    data['idencuesta'] = request.GET['idencuesta']
                    data['cantidad'] = request.GET['cantidad']
                    return render(request, "adm_encuestagrupoestudiantes/addpregunta.html", data)
                except Exception as ex:
                    pass

            if action == 'addrangopregunta':
                try:
                    data['title'] = u'Adicionar Rango Pregunta encuesta'
                    data['form'] = RangoPreguntaEncuestaGrupoEstudiantesForm(initial={})
                    data['idpregunta'] = request.GET['idpregunta']
                    data['cantidad'] = request.GET['cantidad']
                    return render(request, "adm_encuestagrupoestudiantes/addrangopregunta.html", data)
                except Exception as ex:
                    pass

            if action == 'addopcioncuadricula':
                try:
                    data['tipo'] = request.GET['tipo']
                    data['cantidad'] = request.GET['cantidad']
                    data['title'] = u'Adicionar Opción de Fila de Cuadrícula' if int(request.GET['tipo']) == 1 else u'Adicionar Opción de Columna de Cuadrícula'
                    form = OpcionCuadriculaEncuestaGrupoEstudiantesForm(initial={})
                    if int(request.GET['tipo']) == 1:
                        form.sin_secuentia()
                    else:
                        preg = PreguntaEncuestaGrupoEstudiantes.objects.values('encuesta_id').filter(status=True,id=int(request.GET['idpregunta']))
                        form.fields['secuenciapregunta'].queryset=PreguntaEncuestaGrupoEstudiantes.objects.filter(status=True,encuesta__id__in=preg).exclude(id=int(request.GET['idpregunta']))
                    data['form'] = form
                    data['idpregunta'] = request.GET['idpregunta']
                    return render(request, "adm_encuestagrupoestudiantes/addopcioncuadricula.html", data)
                except Exception as ex:
                    pass

            if action == 'addopcionmultiple':
                try:
                    data['tipo'] = request.GET['tipo']
                    data['cantidad'] = request.GET['cantidad']
                    data['title'] = u'Adicionar Opción de respues multiple'
                    data['form'] = OpcionMultipleEncuestaGrupoEstudiantesForm(initial={})
                    data['idpregunta'] = request.GET['idpregunta']
                    return render(request, "adm_encuestagrupoestudiantes/addopcionmultiple.html", data)
                except Exception as ex:
                    pass

            elif action == 'edit':
                try:
                    data['title'] = u'Editar encuesta'
                    encuesta = EncuestaGrupoEstudiantes.objects.get(pk=request.GET['id'])
                    if EncuestaGrupoEstudianteSeguimientoSilabo.objects.filter(
                        encuestagrupoestudiantes_id=request.GET['id']).exists():
                        encuesta_seguimiento = EncuestaGrupoEstudianteSeguimientoSilabo.objects.get(
                            encuestagrupoestudiantes_id=request.GET['id'])
                        categoria = encuesta_seguimiento.categoria
                    else:
                        categoria = False
                    f = EncuestaGrupoEstudiantesForm(initial={'descripcion': encuesta.descripcion,
                                                              'tipoperfil': encuesta.tipoperfil,
                                                              'activo': encuesta.activo,
                                                              'obligatoria': encuesta.obligatoria,
                                                              'leyenda': encuesta.leyenda,
                                                              'categoria': categoria})
                    data['form'] = f
                    data['encuesta'] = encuesta
                    return render(request, "adm_encuestagrupoestudiantes/edit.html", data)
                except Exception as ex:
                    pass

            elif action == 'editpregunta':
                try:
                    data['title'] = u'Editar Pregunta encuesta'
                    pregunta = PreguntaEncuestaGrupoEstudiantes.objects.get(pk=request.GET['id'])
                    f = PreguntaEncuestaGrupoEstudiantesForm(initial={'tipo': pregunta.tipo,
                                                                      'descripcion': pregunta.descripcion,
                                                                      'orden': pregunta.orden,
                                                                      'observacionporno': pregunta.observacionporno,
                                                                      'obligatoria': pregunta.obligatoria,
                                                                      })
                    data['form'] = f
                    data['pregunta'] = pregunta
                    data['cantidad'] = request.GET['cantidad']
                    return render(request, "adm_encuestagrupoestudiantes/editpregunta.html", data)
                except Exception as ex:
                    pass

            elif action == 'editrangopregunta':
                try:
                    data['title'] = u'Editar Rango Pregunta encuesta'
                    data['cantidad'] = request.GET['cantidad']
                    rango = RangoPreguntaEncuestaGrupoEstudiantes.objects.get(pk=request.GET['id'])
                    f = RangoPreguntaEncuestaGrupoEstudiantesForm(initial={'descripcion': rango.descripcion,
                                                                          'orden': rango.orden,
                                                                          'valor': rango.valor})
                    data['form'] = f
                    data['rango'] = rango
                    return render(request, "adm_encuestagrupoestudiantes/editrangopregunta.html", data)
                except Exception as ex:
                    pass

            elif action == 'editopcioncuadricula':
                try:
                    data['tipo'] = request.GET['tipo']
                    data['cantidad'] = request.GET['cantidad']
                    data['idpregunta'] = request.GET['idpregunta']
                    data['title'] = u'Editar Opción de Fila de Cuadrícula' if int(request.GET['tipo']) == 1 else u'Editar Opción de Columna de Cuadrícula'

                    opcion = OpcionCuadriculaEncuestaGrupoEstudiantes.objects.get(pk=request.GET['id'])

                    f = OpcionCuadriculaEncuestaGrupoEstudiantesForm(initial={'descripcion': opcion.descripcion,
                                                                              'orden': opcion.orden,
                                                                              'valor': opcion.valor,
                                                                              'otros': opcion.opcotros,
                                                                              'archivo': opcion.oparchivo,
                                                                              'secuenciapregunta': opcion.secuenciapregunta,
                                                                              })
                    if int(request.GET['tipo']) == 1:
                        f.sin_secuentia()
                    else:
                        preg = PreguntaEncuestaGrupoEstudiantes.objects.values('encuesta_id').filter(status=True,
                                                                                                     id=int(request.GET[
                                                                                                                'idpregunta']))
                        f.fields['secuenciapregunta'].queryset = PreguntaEncuestaGrupoEstudiantes.objects.filter(
                            status=True, encuesta__id__in=preg).exclude(id=int(request.GET['idpregunta']))

                    data['form'] = f
                    data['opcion'] = opcion
                    return render(request, "adm_encuestagrupoestudiantes/editopcioncuadricula.html", data)
                except Exception as ex:
                    pass

            elif action == 'editopcionmultiple':
                try:
                    data['tipo'] = request.GET['tipo']
                    data['cantidad'] = request.GET['cantidad']
                    data['idpregunta'] = request.GET['idpregunta']
                    data['title'] = u'Editar Opción de respuesta múltiple'
                    opcion = OpcionMultipleEncuestaGrupoEstudiantes.objects.get(pk=request.GET['id'])

                    f = OpcionMultipleEncuestaGrupoEstudiantesForm(initial={'descripcion': opcion.descripcion,
                                                                              'orden': opcion.orden,
                                                                              'valor': opcion.valor,
                                                                              'otros':opcion.opcotros})
                    data['form'] = f
                    data['opcion'] = opcion
                    return render(request, "adm_encuestagrupoestudiantes/editopcionmultiple.html", data)
                except Exception as ex:
                    pass

            elif action == 'delete':
                try:
                    data['title'] = u'Eliminar encuesta'
                    data['encuesta'] = EncuestaGrupoEstudiantes.objects.get(pk=request.GET['id'])
                    return render(request, "adm_encuestagrupoestudiantes/delete.html", data)
                except Exception as ex:
                    pass

            elif action == 'deletepregunta':
                try:
                    data['title'] = u'Eliminar pregunta encuesta'
                    data['pregunta'] = PreguntaEncuestaGrupoEstudiantes.objects.get(pk=request.GET['id'])
                    return render(request, "adm_encuestagrupoestudiantes/deletepregunta.html", data)
                except Exception as ex:
                    pass

            elif action == 'deleterangopregunta':
                try:
                    data['title'] = u'Eliminar rango pregunta encuesta'
                    data['rango'] = RangoPreguntaEncuestaGrupoEstudiantes.objects.get(pk=request.GET['id'])
                    return render(request, "adm_encuestagrupoestudiantes/deleterangopregunta.html", data)
                except Exception as ex:
                    pass

            elif action == 'deleteopcioncuadricula':
                try:
                    data['title'] = u'Eliminar opción de respuesta multiple'
                    data['cantidad'] = request.GET['cantidad']
                    data['opcion'] = OpcionCuadriculaEncuestaGrupoEstudiantes.objects.get(pk=request.GET['id'])
                    return render(request, "adm_encuestagrupoestudiantes/deleteopcioncuadricula.html", data)
                except Exception as ex:
                    pass

            elif action == 'deleteopcionmultiple':
                try:
                    data['title'] = u'Eliminar opción de respuesta multiple'
                    data['cantidad'] = request.GET['cantidad']
                    data['opcion'] = OpcionMultipleEncuestaGrupoEstudiantes.objects.get(pk=request.GET['id'])
                    return render(request, "adm_encuestagrupoestudiantes/deleteopcionmultiple.html", data)
                except Exception as ex:
                    pass

            if action == 'preguntas':
                try:
                    data['title'] = u'Preguntas Encuesta'
                    data['idencuesta'] = request.GET['id']
                    preguntas = PreguntaEncuestaGrupoEstudiantes.objects.filter(status=True, encuesta_id=int(request.GET['id'])).order_by('orden')
                    data['preguntas'] = preguntas
                    data['cantidad'] = int(request.GET['cantidad'])
                    return render(request, "adm_encuestagrupoestudiantes/preguntas.html", data)
                except Exception as ex:
                    pass

            if action == 'duplicarconenidopregunta':
                try:
                    data['title'] = u'Duplicar contenido pregunta'
                    pk = int(request.GET['id'])
                    ePreguntaEncuestaGrupoEstudiantes = PreguntaEncuestaGrupoEstudiantes.objects.get(pk=pk)
                    data['action'] = action
                    data['id'] = ePreguntaEncuestaGrupoEstudiantes.pk
                    form = DuplicarContenidoPreguntaForm()
                    form.load_pregunta(ePreguntaEncuestaGrupoEstudiantes.pk,ePreguntaEncuestaGrupoEstudiantes.encuesta_id,ePreguntaEncuestaGrupoEstudiantes.tipo)
                    data['form'] = form
                    template = get_template("adm_configuracionpropuesta/modal/formmodal.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            if action == 'duplicarcontenidoencuesta':
                try:
                    data['title'] = u'Duplicar contenido encuesta'
                    pk = int(request.GET['id'])
                    eEncuestaGrupoEstudiantes = EncuestaGrupoEstudiantes.objects.get(pk=pk)
                    data['action'] = action
                    data['id'] = eEncuestaGrupoEstudiantes.pk
                    form = DuplicarContenidoEncuestaForm()
                    data['form'] = form
                    template = get_template("adm_configuracionpropuesta/modal/formmodal.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            if action == 'rangopregunta':
                try:
                    data['title'] = u'Rango Preguntas Encuesta'
                    data['pregunta'] = PreguntaEncuestaGrupoEstudiantes.objects.get(pk=int(request.GET['id']))
                    rangos = RangoPreguntaEncuestaGrupoEstudiantes.objects.filter(status=True, pregunta_id=int(request.GET['id'])).order_by('orden')
                    data['rangos'] = rangos
                    data['cantidad'] = request.GET['cantidad']
                    return render(request, "adm_encuestagrupoestudiantes/rangopregunta.html", data)
                except Exception as ex:
                    pass

            if action == 'opcionescuadricula':
                try:
                    data['title'] = u'Opciones de la cuadrícula'
                    data['pregunta'] = pregunta = PreguntaEncuestaGrupoEstudiantes.objects.get(pk=int(request.GET['id']))

                    filas = pregunta.opcioncuadriculaencuestagrupoestudiantes_set.filter(tipoopcion=1, status=True).order_by('orden')
                    columnas = pregunta.opcioncuadriculaencuestagrupoestudiantes_set.filter(tipoopcion=2, status=True).order_by('orden')
                    data['filas'] = filas
                    data['columnas'] = columnas
                    data['cantidad'] = request.GET['cantidad']

                    return render(request, "adm_encuestagrupoestudiantes/opcioncuadricula.html", data)
                except Exception as ex:
                    pass

            if action == 'opcionesmultiples':
                try:
                    data['title'] = u'Opciones de respuesta múltiple'
                    data['pregunta'] = pregunta = PreguntaEncuestaGrupoEstudiantes.objects.get(pk=int(request.GET['id']))
                    data['listadomultiple'] = pregunta.opcionmultipleencuestagrupoestudiantes_set.filter(status=True).order_by('orden')
                    data['cantidad'] = request.GET['cantidad']
                    return render(request, "adm_encuestagrupoestudiantes/opcionmultiple.html", data)
                except Exception as ex:
                    pass

            elif action == 'resultado':
                try:
                    encuesta = EncuestaGrupoEstudiantes.objects.get(pk=int(request.GET['id']))
                    # reporte_encuesta_grupo_estudiante_background()
                    noti = Notificacion(cuerpo='Generación de reporte de excel en progreso',
                                        titulo='Resultados de Encuesta', destinatario=persona,
                                        url='',
                                        prioridad=1, app_label='SGA',
                                        fecha_hora_visible=datetime.now() + timedelta(days=1), tipo=2,
                                        en_proceso=True)
                    noti.save(request)
                    reporte_encuesta_grupo_estudiante_background(request=request, notiid=noti.pk, encuesta=encuesta).start()
                    return JsonResponse({"result": True,
                                         "mensaje": u"El reporte se está realizando. Verifique su apartado de notificaciones después de unos minutos.",
                                         "btn_notificaciones": traerNotificaciones(request, data, persona)})
                    # __author__ = 'Unemi'
                    # style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
                    # style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
                    # style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
                    # title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                    # style1 = easyxf(num_format_str='D-MMM-YY')
                    # font_style = XFStyle()
                    # font_style.font.bold = True
                    # font_style2 = XFStyle()
                    # font_style2.font.bold = False
                    # wb = Workbook(encoding='utf-8')
                    # ws = wb.add_sheet('datos')
                    # ws.write_merge(0, 0, 0, 5, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                    # response = HttpResponse(content_type="application/ms-excel")
                    # response['Content-Disposition'] = 'attachment; filename=datos_' + random.randint(1, 10000).__str__() + '.xls'
                    # encuesta = EncuestaGrupoEstudiantes.objects.get(pk=int(request.GET['id']))
                    # preguntas = encuesta.preguntaencuestagrupoestudiantes_set.filter(status=True).order_by('orden')
                    # columns = [
                    #     (u"Nº.", 2000),
                    #     (u"CÉDULA", 3000),
                    #     (u"ENCUESTADO", 9000),
                    # ]
                    # desde = int(request.GET['desde']) if 'desde' in request.GET else None
                    # hasta = int(request.GET['hasta']) if 'hasta' in request.GET else None
                    # if encuesta.tipoperfil == 1:
                    #     columns.append((u'CARRERA', 9000), )
                    # if encuesta.pk == 14:
                    #     columns.append((u'Tipo de relación laboral', 9000), )
                    #     columns.append((u'Tiempo de dedicación', 9000), )
                    #     columns.append((u'Si es docente titular a qué categoría académica pertenece', 9000), )
                    #     # columns.append((u'MODALIDAD CONTRATACIÓN', 9000), )
                    #     columns.append((u'A qué Facultad pertenece', 9000), )
                    #     columns.append((u'Las carreras en las que imparte docencia actualmente ¿De qué modalidad son? ', 9000), )
                    #     # columns.append((u'MODALIDAD DE LA CARRERA QUE DESEARÍA TRABAJAR EN EL SEMESTRE 1S 2022', 9000), )
                    # for x in preguntas:
                    #     columns.append((str(x.orden) + ") " + x.descripcion, 6000), )
                    #     if x.tipo == 1:
                    #         if not x.esta_vacia():
                    #             columns.append((str(x.orden) + ") " + x.observacionporno, 6000), )
                    # row_num = 3
                    # for col_num in range(len(columns)):
                    #     ws.write(row_num, col_num, columns[col_num][0], font_style)
                    #     ws.col(col_num).width = columns[col_num][1]
                    # row_num = 4
                    # i = 0
                    # if encuesta.tipoperfil == 1:
                    #     datos = encuesta.inscripcionencuestagrupoestudiantes_set.filter(status=True, respondio=True).order_by('inscripcion__persona__apellido1')
                    # else:
                    #     datos = encuesta.inscripcionencuestagrupoestudiantes_set.filter(status=True, respondio=True).order_by('profesor__persona__apellido1')
                    # cout_register = datos.count()
                    # register_start = 0
                    # limit = 0
                    # for inscripcion1 in datos:
                    #     row_num += limit
                    #     i += 1
                    #     limit = 0
                    #     if desde is not None and hasta is not None:
                    #         if register_start>=desde and register_start<=hasta:
                    #             ws.write(row_num, 0, i, font_style2)
                    #             if encuesta.tipoperfil == 1:
                    #                 ws.write(row_num, 1, inscripcion1.inscripcion.persona.cedula, font_style2)
                    #                 ws.write(row_num, 2, inscripcion1.inscripcion.persona.nombre_completo_inverso(), font_style2)
                    #             else:
                    #                 ws.write(row_num, 1, inscripcion1.profesor.persona.cedula, font_style2)
                    #                 ws.write(row_num, 2, inscripcion1.profesor.persona.nombre_completo_inverso(), font_style2)
                    #             c = 3
                    #             for x in preguntas:
                    #                 respuesta = inscripcion1.respuestapreguntaencuestagrupoestudiantes_set.get(status=True, pregunta=x) if inscripcion1.respuestapreguntaencuestagrupoestudiantes_set.filter(status=True, pregunta=x).exists() else None
                    #                 if x.tipo == 2:
                    #                     if respuesta is not None:
                    #                         rango = RangoPreguntaEncuestaGrupoEstudiantes.objects.get(pk=int(respuesta.respuesta))
                    #                         ws.write(row_num, c, rango.descripcion, font_style2)
                    #                     else:
                    #                         ws.write(row_num, c, '', font_style2)
                    #                     c += 1
                    #                 else:
                    #                     if respuesta is not None:
                    #                         ws.write(row_num, c, respuesta.respuesta, font_style2)
                    #                     else:
                    #                         ws.write(row_num, c, '', font_style2)
                    #                     c += 1
                    #                     if x.tipo == 1:
                    #                         if not x.esta_vacia():
                    #                             ws.write(row_num, c, respuesta.respuestaporno, font_style2)
                    #                             c += 1
                    #             row_num += 1
                    #             register_start += 1
                    #         else:
                    #             register_start += 1
                    #     else:
                    #         ws.write(row_num, 0, i, font_style2)
                    #         if encuesta.tipoperfil == 1:
                    #             ws.write(row_num, 1, inscripcion1.inscripcion.persona.cedula, font_style2)
                    #             ws.write(row_num, 2, inscripcion1.inscripcion.persona.nombre_completo_inverso(),
                    #                      font_style2)
                    #         else:
                    #             ws.write(row_num, 1, inscripcion1.profesor.persona.cedula, font_style2)
                    #             ws.write(row_num, 2, inscripcion1.profesor.persona.nombre_completo_inverso(),
                    #                      font_style2)
                    #         c = 3
                    #         if inscripcion1.inscripcion:
                    #             ws.write(row_num, c, inscripcion1.inscripcion.carrera.__str__(), font_style2) if not inscripcion1.inscripcion.carrera == None else ' '
                    #             c+=1
                    #
                    #         if encuesta.pk == 14:
                    #             dt = ProfesorDistributivoHoras.objects.filter(status=True, periodo=119, profesor_id=inscripcion1.profesor.id).first()
                    #             ws.write(row_num, c, dt.nivelcategoria.nombre if dt is not None else '', font_style2)
                    #             c += 1
                    #             ws.write(row_num, c, dt.dedicacion.nombre if dt is not None else '', font_style2)
                    #             c += 1
                    #             ws.write(row_num, c, dt.categoria.nombre if dt is not None and dt.nivelcategoria.id == 1 else '', font_style2)
                    #             c += 1
                    #             ws.write(row_num, c, dt.coordinacion.nombre if dt is not None and dt.coordinacion is not None else '', font_style2)
                    #             c += 1
                    #             w = 0
                    #             for m in inscripcion1.profesor.mis_materias(119).values_list('materia__nivel__modalidad__nombre', flat=True).distinct('materia__nivel__modalidad__nombre'):
                    #                 ws.write(row_num + w, c,  str(m), font_style2)
                    #                 w += 1
                    #             if limit < w and w > 0:
                    #                 limit = w - 1
                    #
                    #             c += 1
                    #
                    #         for x in preguntas:
                    #             respuesta = inscripcion1.respuestapreguntaencuestagrupoestudiantes_set.get(status=True,
                    #                                                                                        pregunta=x) if inscripcion1.respuestapreguntaencuestagrupoestudiantes_set.filter(
                    #                 status=True, pregunta=x).exists() else None
                    #             if x.tipo == 2:
                    #                 if respuesta is not None:
                    #                     rango = RangoPreguntaEncuestaGrupoEstudiantes.objects.get(
                    #                         pk=int(respuesta.respuesta))
                    #                     ws.write(row_num, c, rango.descripcion, font_style2)
                    #                 else:
                    #                     ws.write(row_num, c, '', font_style2)
                    #                 c += 1
                    #             elif x.tipo == 5:
                    #                 respuesta = inscripcion1.respuestacuadriculaencuestagrupoestudiantes_set.get(status=True, pregunta=x) if inscripcion1.respuestacuadriculaencuestagrupoestudiantes_set.filter(status=True, pregunta=x).exists() else None
                    #                 if respuesta is not None:
                    #                     try:
                    #                         int(respuesta.respuesta)
                    #                         resp = OpcionCuadriculaEncuestaGrupoEstudiantes.objects.get(status=True, pregunta=x, valor=respuesta.respuesta, tipoopcion=2).descripcion
                    #                     except ValueError:
                    #                         resp = respuesta.respuesta
                    #
                    #                     ws.write(row_num, c, resp, font_style2)
                    #                 else:
                    #                     ws.write(row_num, c, '', font_style2)
                    #                 c += 1
                    #             elif x.tipo == 6:
                    #                 respuesta = inscripcion1.respuestamultipleencuestagrupoestudiantes_set.filter(status=True, pregunta=x) if inscripcion1.respuestamultipleencuestagrupoestudiantes_set.values('id').filter(status=True, pregunta=x).exists() else None
                    #                 if respuesta is not None:
                    #                     w = 0
                    #                     for rmult in respuesta:
                    #                         ws.write(row_num+w, c, rmult.opcionmultiple.descripcion, font_style2)
                    #                         # row_num += 1
                    #                         w += 1
                    #                     if limit < w and w > 0:
                    #                         limit = w-1
                    #                 else:
                    #                     ws.write(row_num, c, '', font_style2)
                    #                 c += 1
                    #             else:
                    #                 if respuesta is not None:
                    #                     ws.write(row_num, c, respuesta.respuesta, font_style2)
                    #                 else:
                    #                     ws.write(row_num, c, '', font_style2)
                    #                 c += 1
                    #                 if x.tipo == 1:
                    #                     if not x.esta_vacia():
                    #                         ws.write(row_num, c, respuesta.respuestaporno, font_style2)
                    #                         c += 1
                    #         row_num += 1
                    # wb.save(response)
                    # return response
                except Exception as ex:
                    print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
                    pass

            elif action == 'importarPoblacion':
                try:

                    data['id_encuesta'] = request.GET['id']
                    data['action'] = 'importarPoblacion'
                    template = get_template("adm_encuestagrupoestudiantes/importarPoblacion.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    transaction.set_rollback(True)
                    translator = Translator()
                    return JsonResponse({"result": "bad", "ex": translator.translate(ex.__str__(), 'es').text, "mensaje": u"Error al eliminar los datos."})

            elif action == 'inscripcionesEncuestas':
                try:
                    data['title'] = u'Inscripciones Encuestas'
                    encuesta = EncuestaGrupoEstudiantes.objects.get(status=True, pk=request.GET['id'])
                    search =request.GET.get('s', '')
                    tipo_perfil = request.GET.get('tipo_perfil', '')
                    ids =request.GET.get('id', '')
                    url_vars = ""
                    filtros = Q(encuesta__id=ids, status=True )

                    if search:
                        url_vars += "&s={}".format(search)
                        if tipo_perfil:
                            if int(tipo_perfil) == 1:
                                filtros = filtros & (Q(inscripcion__persona__nombres__icontains=search) | Q(inscripcion__persona__apellido1__icontains=search) | Q(inscripcion__persona__apellido2__icontains=search)| Q(inscripcion__persona__cedula__icontains=search))
                            if int(tipo_perfil) == 2:
                                filtros = filtros & (Q(profesor__persona__nombres__icontains=search) | Q(profesor__persona__apellido1__icontains=search) | Q(profesor__persona__apellido2__icontains=search )| Q(profesor__persona__cedula__icontains=search ))
                            if int(tipo_perfil) == 3:
                                filtros = filtros & (Q(administrativo__persona__nombres__icontains=search) | Q(administrativo__persona__apellido1__icontains=search ) | Q(administrativo__persona__apellido2__icontains=search )| Q(administrativo__persona__cedula__icontains=search ))



                    inscripciones = InscripcionEncuestaGrupoEstudiantes.objects.filter(filtros).order_by('-id')
                    data["url_vars"] = url_vars
                    paging = MiPaginador(inscripciones, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['titulo_encuesta'] = encuesta.descripcion
                    data['tipo'] = encuesta.get_tipoperfil_display
                    data['tipo_id'] = encuesta.tipoperfil
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['inscripcion_encuesta'] = page.object_list
                    data['encuesta'] = encuesta
                    return render(request, "adm_encuestagrupoestudiantes/listainscripcionesEncuestas.html", data)
                except Exception as ex:
                    pass

            elif action == 'searchPoblacion':
                try:
                    encuesta = EncuestaGrupoEstudiantes.objects.get(pk=request.GET['ide'])
                    filtros = None
                    search = None
                    tipoperfil = 0
                    txt_filter = request.GET['sSearch'] if request.GET['sSearch'] else ''
                    limit = int(request.GET['iDisplayLength']) if request.GET['iDisplayLength'] else 25
                    offset = int(request.GET['iDisplayStart']) if request.GET['iDisplayStart'] else 0
                    aaData = []
                    tCount = 0
                    if txt_filter:
                        search = txt_filter.strip()
                    poblacion = InscripcionEncuestaGrupoEstudiantes.objects.filter(encuesta=encuesta)
                    if encuesta.tipoperfil == 1:
                        title = u'Población de inscripciones'
                        registros = Inscripcion.objects.filter(status=True).exclude(pk__in=poblacion.values_list("inscripcion_id", flat=True))
                        tipoperfil = encuesta.tipoperfil
                        if search:
                            ss = search.split(' ')
                            if len(ss) == 1:
                                registros = registros.filter(Q(persona__nombres__icontains=search) |
                                                             Q(persona__cedula__icontains=search) |
                                                             Q(persona__pasaporte__icontains=search) |
                                                             Q(persona__apellido1__icontains=search) |
                                                             Q(persona__apellido2__icontains=search))
                            else:
                                registros = registros.filter(Q(persona__nombres__icontains=search) |
                                                             Q(persona__cedula__icontains=search) |
                                                             Q(persona__pasaporte__icontains=search) |
                                                             Q(persona__apellido1__icontains=ss[0]) |
                                                             Q(persona__apellido2__icontains=ss[1]))
                    elif encuesta.tipoperfil == 2:
                        title = u'Población de docentes'
                        registros = Profesor.objects.filter(status=True).exclude(pk__in=poblacion.values_list("profesor_id", flat=True))
                        tipoperfil = encuesta.tipoperfil
                        if search:
                            ss = search.split(' ')
                            if len(ss) == 1:
                                registros = registros.filter(Q(persona__nombres__icontains=search) |
                                                             Q(persona__cedula__icontains=search) |
                                                             Q(persona__pasaporte__icontains=search) |
                                                             Q(persona__apellido1__icontains=search) |
                                                             Q(persona__apellido2__icontains=search))
                            else:
                                registros = registros.filter(Q(persona__nombres__icontains=search) |
                                                             Q(persona__cedula__icontains=search) |
                                                             Q(persona__pasaporte__icontains=search) |
                                                             Q(persona__apellido1__icontains=ss[0]) |
                                                             Q(persona__apellido2__icontains=ss[1]))
                    elif encuesta.tipoperfil == 3:
                        title = u'Población de administrativios'
                        registros = Administrativo.objects.filter(status=True).exclude(pk__in=poblacion.values_list("administrativo_id", flat=True))
                        tipoperfil = encuesta.tipoperfil
                        if search:
                            ss = search.split(' ')
                            if len(ss) == 1:
                                registros = registros.filter(Q(persona__nombres__icontains=search) |
                                                             Q(persona__cedula__icontains=search) |
                                                             Q(persona__pasaporte__icontains=search) |
                                                             Q(persona__apellido1__icontains=search) |
                                                             Q(persona__apellido2__icontains=search))
                            else:
                                registros = registros.filter(Q(persona__nombres__icontains=search) |
                                                             Q(persona__cedula__icontains=search) |
                                                             Q(persona__pasaporte__icontains=search) |
                                                             Q(persona__apellido1__icontains=ss[0]) |
                                                             Q(persona__apellido2__icontains=ss[1]))
                    else:
                        title = u'Población general'
                        registros = Persona.objects.filter(status=True)
                        tipoperfil = 0
                        if search:
                            ss = search.split(' ')
                            if len(ss) == 1:
                                registros = registros.filter(Q(nombres__icontains=search) |
                                                             Q(cedula__icontains=search) |
                                                             Q(pasaporte__icontains=search) |
                                                             Q(apellido1__icontains=search) |
                                                             Q(apellido2__icontains=search))
                            else:
                                registros = registros.filter(Q(nombres__icontains=search) |
                                                             Q(cedula__icontains=search) |
                                                             Q(pasaporte__icontains=search) |
                                                             Q(apellido1__icontains=ss[0]) |
                                                             Q(apellido2__icontains=ss[1]))
                    data['title'] = title
                    tCount = registros.count()
                    if offset == 0:
                        rows = registros[offset:limit]
                    else:
                        rows = registros[offset:offset + limit]
                    aaData = []
                    for row in rows:
                        detalle = {}
                        if encuesta.tipoperfil in [1, 2, 3]:
                            documento = row.persona.documento()
                            nombre_completo = row.persona.nombre_completo_inverso()
                            sexo = row.persona.sexo.nombre if row.persona.sexo_id else ''
                            if encuesta.tipoperfil == 1:
                                detalle = {"carrera": row.carrera.__str__(),
                                           "tipoperfil": encuesta.tipoperfil}
                            elif encuesta.tipoperfil == 2:
                                eDistributivoPersonas = DistributivoPersona.objects.filter(persona=row.persona, status=True, regimenlaboral_id__in=[1,4], estadopuesto_id=1)
                                if eDistributivoPersonas.values("id").exists():
                                    eDistributivoPersona = eDistributivoPersonas.first()
                                    detalle = {"regimenlaboral": eDistributivoPersona.regimenlaboral.descripcion,
                                               "denominacionpuesto": eDistributivoPersona.denominacionpuesto.descripcion,
                                               "tipoperfil": encuesta.tipoperfil}
                            elif encuesta.tipoperfil == 3:
                                eDistributivoPersonas = DistributivoPersona.objects.filter(persona=row.persona, status=True, regimenlaboral_id=2, estadopuesto_id=1)
                                if eDistributivoPersonas.values("id").exists():
                                    eDistributivoPersona = eDistributivoPersonas.first()
                                    detalle = {"regimenlaboral": eDistributivoPersona.regimenlaboral.descripcion,
                                               "denominacionpuesto": eDistributivoPersona.denominacionpuesto.descripcion,
                                               "tipoperfil": encuesta.tipoperfil}

                        else:
                            documento = row.documento()
                            nombre_completo = row.nombre_completo_inverso()
                            sexo = row.sexo.nombre if row.sexo_id else ''
                        aaData.append([row.id,
                                       documento,
                                       nombre_completo,
                                       sexo,
                                       detalle,
                                       {"id": row.id,
                                        "ide": encuesta.id,
                                        "tipoperfil": tipoperfil,
                                        "nombre_completo": nombre_completo,
                                        }
                                       ])
                    data['encuesta'] = encuesta
                    # template = get_template("adm_encuestagrupoestudiantes/listadopoblacion.html")
                    # json_content = template.render(data)
                    return JsonResponse({"result": "ok", "data": aaData, "iTotalRecords": tCount, "iTotalDisplayRecords": tCount, 'title': title})
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos. %s" % ex.__str__()})
            return HttpResponseRedirect(request.path)
        else:
            try:
                search = None
                ids = None
                encuestas = EncuestaGrupoEstudiantes.objects.filter(status=True).order_by('-id')
                if 's' in request.GET:
                    search = request.GET['s']
                    ss = search.split(' ')
                    if len(ss) == 2:
                        encuestas = encuestas.filter(Q(descripcion__icontains=ss[0]) & Q(descripcion__icontains=ss[1]))
                    if len(ss) == 3:
                        encuestas = encuestas.filter(Q(descripcion__icontains=ss[0]) & Q(descripcion__icontains=ss[1]) & Q(descripcion__icontains=ss[2]))
                    if len(ss) == 4:
                        encuestas = encuestas.filter(Q(descripcion__icontains=ss[0]) & Q(descripcion__icontains=ss[1]) & Q(descripcion__icontains=ss[2]) & Q(descripcion__icontains=ss[3]))
                    else:
                        encuestas = encuestas.filter(descripcion__icontains=search, status=True)
                elif 'id' in request.GET:
                    ids = request.GET['id']
                    encuestas = encuestas.filter(id=ids)
                paging = MiPaginador(encuestas, 25)
                p = 1
                try:
                    paginasesion = 1
                    if 'paginador' in request.session:
                        paginasesion = int(request.session['paginador'])
                    if 'page' in request.GET:
                        p = int(request.GET['page'])
                    else:
                        p = paginasesion
                    try:
                        page = paging.page(p)
                    except:
                        p = 1
                    page = paging.page(p)
                except:
                    page = paging.page(p)
                request.session['paginador'] = p
                data['paging'] = paging
                data['rangospaging'] = paging.rangos_paginado(p)
                data['page'] = page
                data['search'] = search if search else ""
                data['ids'] = ids if ids else ""
                data['encuestas'] = page.object_list
                return render(request, "adm_encuestagrupoestudiantes/view.html", data)
            except Exception as ex:
                return HttpResponseRedirect(f"/?info{ex.__str__()}")
