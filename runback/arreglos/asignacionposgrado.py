# coding=utf-8
#!/usr/bin/env python

import os
import sys


# SITE_ROOT = os.path.dirname(os.path.realpath(__file__))
import xlrd
from openpyxl import load_workbook

YOUR_PATH = os.path.dirname(os.path.realpath(__file__))
# print(f"YOUR_PATH: {YOUR_PATH}")
SITE_ROOT = os.path.dirname(os.path.dirname(YOUR_PATH))
SITE_ROOT = os.path.join(SITE_ROOT, '')
# print(f"SITE_ROOT: {SITE_ROOT}")
your_djangoproject_home = os.path.split(SITE_ROOT)[0]
# print(f"your_djangoproject_home: {your_djangoproject_home}")
sys.path.append(your_djangoproject_home)

from django.core.wsgi import get_wsgi_application
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "settings")
application = get_wsgi_application()
from datetime import datetime, timedelta
from django.db import transaction
from sga.models import Materia, Clase, Leccion, LeccionGrupo, AsistenciaLeccion
from sga.funciones import variable_valor, validar_ldap_reseteo, convertir_hora
from sga.models import *
from sagest.models import *
from moodle.models import UserAuth
from datetime import datetime, timedelta
from django.db import transaction
from sga.models import Materia, Clase, Leccion, LeccionGrupo, AsistenciaLeccion, CamposTitulosPostulacion
from sga.funciones import variable_valor
# from sga.models import *
# from sagest.models import *
import xlrd
from postulate.models import Partida, PersonaAplicarPartida, PersonaFormacionAcademicoPartida, PartidaTribunal
from sga.models import *
from sagest.models import *
from django.db.models import Sum, F, FloatField
from django.db.models.functions import Coalesce
from settings import MEDIA_ROOT, BASE_DIR, CALCULO_POR_CREDITO
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, colors
from django.http import HttpResponse
from sga.models import Modulo
from gdocumental.models import *
from bd.models import *


def asigposgrado(idmatricula, turno):
    try:
        matricula = Matricula.objects.get(pk=int(idmatricula))
        if matricula.inscripcion.bloqueomatricula:
            print('Matricula Bloqueada')
            return False
        materia = Materia.objects.get(pk=int(turno))
        profesormateria = None
        grupoprofesormateria = None
        if matricula.inscripcion.existe_en_malla(materia.asignatura) and not matricula.inscripcion.puede_tomar_materia(materia.asignatura):
            print('No puede tomar esta materia por tener precedencias')
            return False
        if matricula.inscripcion.existe_en_modulos(materia.asignatura) and not matricula.inscripcion.puede_tomar_materia_modulo(materia.asignatura):
            print('No puede tomar esta materia por tener precedencias')
            return False
        if not materia.modeloevaluativo:
            print('No tiene modelo evaluativo en el distributivo de la asignatura')
            return False
        matriculacupoadicional = False
        if matricula.materiaasignada_set.values('id').filter(materia=materia, status=True).exists():
            print('Ya se encuentra matriculado en esta materia')
            return False
        elif matricula.materiaasignada_set.values('id').filter(materia=materia, status=False).exists():
            materiaasig = MateriaAsignada.objects.get(matricula=matricula, materia=materia, status=False)
            materiaasig.status = True
            materiaasig.save()
            print(f'Materia Asignada {materiaasig}')
            return True
        if materia.inglesepunemi:
            materiaasignada = MateriaAsignada(matricula=matricula,
                                              materia=materia,
                                              notafinal=0,
                                              asistenciafinal=100,
                                              cerrado=False,
                                              automatricula=True,
                                              importa_nota=True,
                                              observaciones='',
                                              estado_id=NOTA_ESTADO_EN_CURSO)
            materiaasignada.save()
            matriculas = matricula.inscripcion.historicorecordacademico_set.filter(asignatura=materia.asignatura).count() + 1
            if matriculas > 1 or matricula.inscripcion.persona.tiene_otro_titulo(matricula.inscripcion):
                matricula.nuevo_calculo_matricula_ingles(materiaasignada)
        else:
            if matricula.inscripcion.carrera.modalidad == 3:
                materiaasignada = MateriaAsignada(matricula=matricula,
                                                  materia=materia,
                                                  notafinal=0,
                                                  sinasistencia=True,
                                                  asistenciafinal=100,
                                                  cerrado=False,
                                                  observaciones='',
                                                  estado_id=NOTA_ESTADO_EN_CURSO)
            else:
                materiaasignada = MateriaAsignada(matricula=matricula,
                                                  materia=materia,
                                                  notafinal=0,
                                                  asistenciafinal=0,
                                                  cerrado=False,
                                                  observaciones='',
                                                  estado_id=NOTA_ESTADO_EN_CURSO)
            materiaasignada.save()
        if matriculacupoadicional:
            materia.totalmatriculadocupoadicional += 1
            materia.cupo += 1
            materia.save()
            print(u'Estudiante matriculado en cupo adicional materia: %s - estudiante: %s y se aumento un cupo en materia' % (materia, matricula))
        matricula.actualizar_horas_creditos()
        materiaasignada.matriculas = materiaasignada.cantidad_matriculas()
        materiaasignada.asistencias()
        materiaasignada.evaluacion()
        materiaasignada.mis_planificaciones()
        materiaasignada.save()
        if matricula.nivel.nivelgrado:
            pass
        else:
            if datetime.now().date() < matricula.nivel.periodo.inicio_agregacion:
                # AGREGACION DE MATERIAS EN MATRICULACION REGULAR SIN REALIZAR PAGOS
                materiaasignada.save()
                print(u'Adiciono materia: %s' % materiaasignada)
                    # matricula.calcular_rubros_matricula(cobro)
            elif matricula.nivel.periodo.fecha_agregaciones():
                # AGREGACION DE MATERIAS EN FECHAS DE AGREGACIONES
                pers_ = Persona.objects.get(pk=27604)
                registro = AgregacionEliminacionMaterias(matricula=matricula,
                                                         agregacion=True,
                                                         asignatura=materiaasignada.materia.asignatura,
                                                         responsable=pers_,
                                                         fecha=datetime.now().date(),
                                                         creditos=materiaasignada.materia.creditos,
                                                         nivelmalla=materiaasignada.materia.nivel.nivelmalla if materiaasignada.materia.nivel.nivelmalla else None,
                                                         matriculas=materiaasignada.matriculas)
                registro.save()
                print(u'Adiciono materia: %s' % materiaasignada)
            else:
                # AGREGACION DE MATERIAS TERMINADA LAS AGREGACIONES
                if materia.asignatura.modulo:
                    pers_ = Persona.objects.get(pk=27604)
                    registro = AgregacionEliminacionMaterias(matricula=matricula,
                                                             agregacion=True,
                                                             asignatura=materiaasignada.materia.asignatura,
                                                             responsable=pers_,
                                                             fecha=datetime.now().date(),
                                                             creditos=materiaasignada.materia.creditos,
                                                             nivelmalla=materiaasignada.materia.nivel.nivelmalla if materiaasignada.materia.nivel.nivelmalla else None,
                                                             matriculas=materiaasignada.matriculas)
                    registro.save()
                    print(u'Adiciono materia: %s' % materiaasignada)
                    return True
    except Exception as ex:
        print(ex)
        return False

archivo_ = 'posgrado_matri_masivo'
url_archivo = "{}/media/{}.xlsx".format(SITE_STORAGE, archivo_)
workbook = load_workbook(filename=url_archivo, read_only=False)
sheet = workbook[workbook.sheetnames[0]]
linea = 1
col_matricula = 3
totalmatri = 0
for rowx in range(2, sheet.max_row + 1):
    if True:
        matricula_ = sheet.cell(row=rowx,column=col_matricula).value
        m1 = sheet.cell(row=rowx,column=4).value
        m2 = sheet.cell(row=rowx,column=5).value
        m3 = sheet.cell(row=rowx,column=6).value
        m4 = sheet.cell(row=rowx,column=7).value
        m5 = sheet.cell(row=rowx,column=8).value
        m6 = sheet.cell(row=rowx,column=9).value
        m7 = sheet.cell(row=rowx,column=10).value
        m8 = sheet.cell(row=rowx,column=11).value
        m9 = sheet.cell(row=rowx,column=12).value
        m10 = sheet.cell(row=rowx,column=13).value
        totalmatri += 1
        print(matricula_, m1, m2, m3, m4, m5, m6, m7, m8, m9, m10)
        asigposgrado(matricula_, m1)
        asigposgrado(matricula_, m2)
        asigposgrado(matricula_, m3)
        asigposgrado(matricula_, m4)
        asigposgrado(matricula_, m5)
        asigposgrado(matricula_, m6)
        asigposgrado(matricula_, m7)
        asigposgrado(matricula_, m8)
        asigposgrado(matricula_, m9)
        asigposgrado(matricula_, m10)
print(f'Matriculados: {totalmatri}')
