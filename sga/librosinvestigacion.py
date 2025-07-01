# -*- coding: UTF-8 -*-
import os
import random
import zipfile
from datetime import datetime
import time as ET

import openpyxl
from django.core.files.base import ContentFile
from django.db.models.functions import ExtractYear
from django.template.loader import get_template
import xlwt
from xlwt import *
from django.contrib.auth.decorators import login_required
from django.db import transaction, connection, connections
from django.db.models.query_utils import Q
from django.http import HttpResponseRedirect, HttpResponse, JsonResponse
from django.shortcuts import render
from django.template import Context

from decorators import secure_module, last_access
from investigacion.funciones import coordinador_investigacion, notificar_revision_solicitud_produccion_cientifica, codigo_publicacion_valido
from sagest.commonviews import obtener_estados_solicitud, obtener_estado_solicitud
from sagest.models import SolicitudPublicacion, ParticipanteSolicitudPublicacion
from settings import SITE_STORAGE
from sga.commonviews import adduserdata
from sga.forms import EvidenciaForm, LibroInvestigacionForm, \
    ParticipanteProfesorLibroForm, ParticipanteAdministrativoLibroForm, CapituloLibroInvestigacionForm, \
    ParticipanteProfesorCapituloLibroForm, ParticipanteAdministrativoCapituloLibroForm, ParticipanteProfesorPonenciaForm
from sga.funciones import MiPaginador, log, generar_nombre, cuenta_email_disponible, remover_caracteres, \
    remover_caracteres_especiales_unicode, remover_caracteres_tildes_unicode, elimina_tildes, validar_archivo, cuenta_email_disponible_para_envio
from sga.funcionesxhtml2pdf import conviert_html_to_pdf, convert_html_to_pdf
from sga.models import Evidencia, DetalleEvidencias, TIPO_PARTICIPANTE_INSTITUCION, TIPO_PARTICIPANTE, \
    LibroInvestigacion, ParticipanteLibros, CapituloLibroInvestigacion, \
    ParticipanteCapituloLibros, CUENTAS_CORREOS, miinstitucion, Persona
from sga.tasks import conectar_cuenta, send_html_mail
from sga.templatetags.sga_extras import encrypt


@login_required(redirect_field_name='ret', login_url='/loginsga')
@secure_module
@last_access
@transaction.atomic()
def view(request):
    data = {}
    adduserdata(request, data)
    persona = request.session['persona']
    miscarreras = persona.mis_carreras()
    if request.method == 'POST':
        action = request.POST['action']

        if action == 'add':
            try:
                f = LibroInvestigacionForm(request.POST, request.FILES)
                archivoderecho = None

                if 'archivoderecho' in request.FILES:
                    archivo = request.FILES['archivoderecho']
                    descripcionarchivo = 'Documento Derechos de autor'
                    resp = validar_archivo(descripcionarchivo, archivo, ['pdf'], '10MB')
                    if resp['estado'] != "OK":
                        return JsonResponse({"result": "bad", "titulo": "Atención!!!", "mensaje": resp["mensaje"], "showSwal": "True", "swalType": "warning"})

                    archivoderecho = request.FILES['archivoderecho']
                    archivoderecho._name = generar_nombre("derechoautor_", archivoderecho._name)

                if f.is_valid():
                    if not LibroInvestigacion.objects.values('id').filter(nombrelibro=f.cleaned_data['nombrelibro'], status=True).exists():
                        libro = LibroInvestigacion(nombrelibro=f.cleaned_data['nombrelibro'],
                                                   codisbn=f.cleaned_data['codisbn'],
                                                   editorial=f.cleaned_data['editorial'],
                                                   fechapublicacion=f.cleaned_data['fechapublicacion'],
                                                   numeroedicion=f.cleaned_data['numeroedicion'],
                                                   numeropagina=f.cleaned_data['numeropagina'],
                                                   registroderecho=f.cleaned_data['registroderecho'],
                                                   revi_pare=f.cleaned_data['revi_pare'],
                                                   revi_fili=f.cleaned_data['revi_fili'],
                                                   areaconocimiento=f.cleaned_data['areaconocimiento'],
                                                   subareaconocimiento=f.cleaned_data['subareaconocimiento'],
                                                   subareaespecificaconocimiento=f.cleaned_data['subareaespecificaconocimiento'],
                                                   lineainvestigacion=f.cleaned_data['lineainvestigacion'],
                                                   sublineainvestigacion=f.cleaned_data['sublineainvestigacion'],
                                                   provieneproyecto=f.cleaned_data['provieneproyecto'],
                                                   tipoproyecto=f.cleaned_data['tipoproyecto'] if f.cleaned_data['provieneproyecto'] else None,
                                                   pertenecegrupoinv=f.cleaned_data['pertenecegrupoinv'],
                                                   grupoinvestigacion=f.cleaned_data['grupoinvestigacion'],
                                                   accesoabierto=f.cleaned_data['accesoabierto'])
                        libro.save(request)

                        if f.cleaned_data['provieneproyecto']:
                            if int(f.cleaned_data['tipoproyecto']) != 3:
                                libro.proyectointerno = f.cleaned_data['proyectointerno']
                            else:
                                libro.proyectoexterno = f.cleaned_data['proyectoexterno']
                        libro.save(request)

                        if archivoderecho:
                            evidencia = Evidencia.objects.get(pk=16)
                            detalleevidencia = DetalleEvidencias(evidencia=evidencia,
                                                                 libro=libro,
                                                                 descripcion='REGISTRO DERECHOS DE AUTOR',
                                                                 archivo=archivoderecho)
                            detalleevidencia.save(request)

                        log(u'% adicionó libro de investigacion: %s' % (persona, libro), request, "add")
                        return JsonResponse({"result": "ok", "titulo": "Proceso exitoso!!!", "mensaje": u"Registro guardado con éxito", "showSwal": True})
                    else:
                        return JsonResponse({"result": "bad", "titulo": "Atención!!!", "mensaje": u"El título para el tipo de publicación ya ha sido ingresado anteriormente", "showSwal": "True", "swalType": "warning"})
                else:
                    msg = ",".join([k + ': ' + v[0] for k, v in f.errors.items()])
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "titulo": "Error", "mensaje": u"Error al guardar los datos. [%s]" % msg, "showSwal": "True", "swalType": "error"})
            except Exception as ex:
                msg = ex.__str__()
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "titulo": "Error", "mensaje": u"Error al guardar los datos. [%s]" % msg, "showSwal": "True", "swalType": "error"})

        elif action == 'ingresarsolicitudlibro':
            try:
                solicitud = SolicitudPublicacion.objects.get(pk=int(encrypt(request.POST['ids'])))
                f = LibroInvestigacionForm(request.POST)

                archivoderecho = None

                if 'archivoderecho' in request.FILES:
                    archivoderecho = request.FILES['archivoderecho']
                    descripcionarchivo = 'Archivo Documento derechos de autor'

                    # Validar el archivo
                    resp = validar_archivo(descripcionarchivo, archivoderecho, ['PDF'], '10MB')
                    if resp['estado'] != "OK":
                        return JsonResponse({"result": "bad", "mensaje": resp["mensaje"]})

                    archivoderecho._name = generar_nombre("derechoautor", archivoderecho._name)

                if f.is_valid():
                    estadosolicitud = int(request.POST['estadosolicitud'])

                    # Si el estado es VERIFICADO
                    if estadosolicitud == 57:
                        if not LibroInvestigacion.objects.values('id').filter(nombrelibro=f.cleaned_data['nombrelibro'], status=True).exists():
                            libro = LibroInvestigacion(nombrelibro=f.cleaned_data['nombrelibro'],
                                                       codisbn=f.cleaned_data['codisbn'],
                                                       editorial=f.cleaned_data['editorial'],
                                                       fechapublicacion=f.cleaned_data['fechapublicacion'],
                                                       numeroedicion=f.cleaned_data['numeroedicion'],
                                                       numeropagina=f.cleaned_data['numeropagina'],
                                                       registroderecho=f.cleaned_data['registroderecho'],
                                                       revi_pare=f.cleaned_data['revi_pare'],
                                                       revi_fili=f.cleaned_data['revi_fili'],
                                                       areaconocimiento=f.cleaned_data['areaconocimiento'],
                                                       subareaconocimiento=f.cleaned_data['subareaconocimiento'],
                                                       subareaespecificaconocimiento=f.cleaned_data['subareaespecificaconocimiento'],
                                                       lineainvestigacion=f.cleaned_data['lineainvestigacion'],
                                                       sublineainvestigacion=f.cleaned_data['sublineainvestigacion'],
                                                       provieneproyecto=f.cleaned_data['provieneproyecto'],
                                                       tipoproyecto=f.cleaned_data['tipoproyecto'] if f.cleaned_data['provieneproyecto'] else None,
                                                       pertenecegrupoinv=f.cleaned_data['pertenecegrupoinv'],
                                                       grupoinvestigacion=f.cleaned_data['grupoinvestigacion'],
                                                       accesoabierto=f.cleaned_data['accesoabierto'],
                                                       solicitudpublicacion=solicitud)
                            libro.save(request)

                            if f.cleaned_data['provieneproyecto']:
                                if int(f.cleaned_data['tipoproyecto']) != 3:
                                    libro.proyectointerno = f.cleaned_data['proyectointerno']
                                else:
                                    libro.proyectoexterno = f.cleaned_data['proyectoexterno']
                            libro.save(request)

                            solicitud.registrado = True
                            solicitud.estado_id = estadosolicitud
                            solicitud.save(request)

                            # LIBRO
                            evidencia = Evidencia.objects.get(pk=7)
                            new_file = ContentFile(solicitud.archivo.file.read())
                            new_file.name = generar_nombre("libro", solicitud.nombre_archivo_publicacion())

                            detalleevidencia = DetalleEvidencias(evidencia=evidencia,
                                                                 libro=libro,
                                                                 descripcion='LIBRO PUBLICADO',
                                                                 archivo=new_file)
                            detalleevidencia.save(request)

                            # DOCUMENTO DERECHOS DE AUTOR
                            if archivoderecho:
                                evidencia = Evidencia.objects.get(pk=16)
                                detalleevidencia = DetalleEvidencias(evidencia=evidencia,
                                                                     libro=libro,
                                                                     descripcion='REGISTRO DERECHOS DE AUTOR',
                                                                     archivo=archivoderecho)
                                detalleevidencia.save(request)

                            # CERTIFICADO DE PUBLICACIÓN
                            if solicitud.archivocertificado:
                                evidencia = Evidencia.objects.get(pk=28)
                                new_file = ContentFile(solicitud.archivocertificado.file.read())
                                new_file.name = generar_nombre("certificadopublicacion", solicitud.nombre_archivo_carta_aceptacion())

                                detalleevidencia = DetalleEvidencias(evidencia=evidencia,
                                                                     libro=libro,
                                                                     descripcion='CERTIFICADO DE PUBLICACIÓN',
                                                                     archivo=new_file)
                                detalleevidencia.save(request)

                            # MATRIZ REVISION PARES
                            if solicitud.archivocomite:
                                evidencia = Evidencia.objects.get(pk=29)
                                new_file = ContentFile(solicitud.archivocomite.file.read())
                                new_file.name = generar_nombre("revisionpares", solicitud.nombre_archivo_portada_indice())

                                detalleevidencia = DetalleEvidencias(evidencia=evidencia,
                                                                     libro=libro,
                                                                     descripcion='CERTIFICADO O MATRIZ DE REVISIÓN POR PARES',
                                                                     archivo=new_file)
                                detalleevidencia.save(request)

                            # Participantes: Autores/Coautores
                            participantes = ParticipanteSolicitudPublicacion.objects.filter(solicitud=solicitud, status=True).order_by('id')
                            for p in participantes:
                                participantepublicacion = ParticipanteLibros(matrizevidencia_id=6,
                                                                          libros=libro,
                                                                          profesor=p.profesor,
                                                                          tipoparticipante=p.tipo,
                                                                          administrativo=p.administrativo,
                                                                          tipoparticipanteins=p.tipoparticipanteins,
                                                                          inscripcion=p.inscripcion)
                                participantepublicacion.save(request)

                            notificar_revision_solicitud_produccion_cientifica(solicitud)

                            log(u'Adicionó libro de investigación: %s' % libro, request, "add")
                            return JsonResponse({"result": "ok", "titulo": "Proceso exitoso!!!", "mensaje": u"Registro guardado con éxito", "showSwal": True})
                        else:
                            return JsonResponse({"result": "bad", "titulo": "Atención!!!", "mensaje": u"El libro con ese título ya está ingresado", "showSwal": "True", "swalType": "warning"})
                    else:
                        observacion = request.POST['observacionsolicitud'].strip().upper()

                        # Actualiza solicitud
                        solicitud.estado_id = estadosolicitud
                        solicitud.observacion = observacion
                        solicitud.save(request)

                        notificar_revision_solicitud_produccion_cientifica(solicitud)

                        if solicitud.estado.valor == 3:
                            log(u'Registró novedad en solicitud: %s' % solicitud.persona, request, "edit")
                        else:
                            log(u'Rechazó solicitud de artículo: %s' % solicitud.persona, request, "edit")

                        return JsonResponse({"result": "ok", "titulo": "Proceso exitoso!!!", "mensaje": u"Registro guardado con éxito", "showSwal": True})

                else:
                     raise NameError('Error')
            except Exception as ex:
                msg = ex.__str__()
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "titulo": "Error", "mensaje": u"Error al guardar los datos. [%s]" % msg, "showSwal": "True", "swalType": "error"})

        elif action == 'addcapitulo':
            try:
                solicitud = None
                if 'ids' in request.POST:
                    if request.POST['ids'] != '':
                        solicitud = SolicitudPublicacion.objects.get(pk=int(encrypt(request.POST['ids'])))

                f = CapituloLibroInvestigacionForm(request.POST)
                if f.is_valid():
                    estadosolicitud = 0

                    if solicitud:
                        estadosolicitud = int(request.POST['estadosolicitud'])

                        if estadosolicitud != 57:
                            observacion = request.POST['observacionsolicitud'].strip().upper()

                            # Actualiza solicitud
                            solicitud.estado_id = estadosolicitud
                            solicitud.observacion = observacion
                            solicitud.save(request)

                            notificar_revision_solicitud_produccion_cientifica(solicitud)

                            if solicitud.estado.valor == 3:
                                log(u'Registró novedad en solicitud: %s' % solicitud.persona, request, "edit")
                            else:
                                log(u'Rechazó solicitud de capítulo de libro: %s' % solicitud.persona, request, "edit")

                            return JsonResponse({"result": "ok", "titulo": "Proceso exitoso!!!", "mensaje": u"Registro guardado con éxito", "showSwal": True})

                    if not CapituloLibroInvestigacion.objects.values('id').filter(titulocapitulo=f.cleaned_data['titulocapitulo'], status=True).exists():
                        capitulo = CapituloLibroInvestigacion(
                            titulocapitulo=f.cleaned_data['titulocapitulo'],
                            titulolibro=f.cleaned_data['titulolibro'],
                            codisbn=f.cleaned_data['codisbn'],
                            editorcompilador=f.cleaned_data['editorcompilador'],
                            totalcapitulo=f.cleaned_data['totalcapitulo'],
                            filiacion=f.cleaned_data['filiacion'],
                            fechapublicacion=f.cleaned_data['fechapublicacion'],
                            paginas=f.cleaned_data['paginas'],
                            areaconocimiento=f.cleaned_data['areaconocimiento'],
                            subareaconocimiento=f.cleaned_data['subareaconocimiento'],
                            subareaespecificaconocimiento=f.cleaned_data['subareaespecificaconocimiento'],
                            lineainvestigacion=f.cleaned_data['lineainvestigacion'],
                            sublineainvestigacion=f.cleaned_data['sublineainvestigacion'],
                            provieneproyecto=f.cleaned_data['provieneproyecto'],
                            tipoproyecto=f.cleaned_data['tipoproyecto'] if f.cleaned_data['provieneproyecto'] else None,
                            pertenecegrupoinv=f.cleaned_data['pertenecegrupoinv'],
                            grupoinvestigacion=f.cleaned_data['grupoinvestigacion'],
                            revisadopar=f.cleaned_data['revisadopar'],
                            solicitudpublicacion=solicitud
                        )
                        capitulo.save(request)

                        if f.cleaned_data['provieneproyecto']:
                            if int(f.cleaned_data['tipoproyecto']) != 3:
                                capitulo.proyectointerno = f.cleaned_data['proyectointerno']
                            else:
                                capitulo.proyectoexterno = f.cleaned_data['proyectoexterno']
                        capitulo.save(request)

                        if solicitud:
                            solicitud.registrado = True
                            solicitud.estado_id = estadosolicitud
                            solicitud.save(request)

                            # CAPITULO
                            if solicitud.archivo:
                                evidencia = Evidencia.objects.get(pk=12)
                                new_file = ContentFile(solicitud.archivo.file.read())
                                new_file.name = generar_nombre("capitulo", solicitud.nombre_archivo_publicacion())

                                detalleevidencia = DetalleEvidencias(evidencia=evidencia,
                                                                     capitulo=capitulo,
                                                                     descripcion='CAPÍTULO DE LIBRO PUBLICADO',
                                                                     archivo=new_file)
                                detalleevidencia.save(request)

                            # LIBRO
                            if solicitud.archivocertificado:
                                evidencia = Evidencia.objects.get(pk=11)
                                new_file = ContentFile(solicitud.archivocertificado.file.read())
                                new_file.name = generar_nombre("libro", solicitud.nombre_archivo_publicacion())

                                detalleevidencia = DetalleEvidencias(evidencia=evidencia,
                                                                     capitulo=capitulo,
                                                                     descripcion='LIBRO PUBLICADO',
                                                                     archivo=new_file)
                                detalleevidencia.save(request)

                            # CARTA DE ACEPTACION o FICHA CATALOGRAFICA
                            if solicitud.archivoparticipacion:
                                evidencia = Evidencia.objects.get(pk=14)
                                new_file = ContentFile(solicitud.archivoparticipacion.file.read())
                                new_file.name = generar_nombre("certificadopublicacion", solicitud.nombre_archivo_publicacion())

                                detalleevidencia = DetalleEvidencias(evidencia=evidencia,
                                                                     capitulo=capitulo,
                                                                     descripcion='CERTIFICADO DE PUBLICACIÓN',
                                                                     archivo=new_file)
                                detalleevidencia.save(request)

                            # INFORME REVISIÓN PARES
                            if solicitud.archivocomite:
                                evidencia = Evidencia.objects.get(pk=13)
                                new_file = ContentFile(solicitud.archivocomite.file.read())
                                new_file.name = generar_nombre("revisionpares", solicitud.nombre_archivo_publicacion())

                                detalleevidencia = DetalleEvidencias(evidencia=evidencia,
                                                                     capitulo=capitulo,
                                                                     descripcion='CERTIFICADO O MATRIZ DE REVISIÓN POR PARES',
                                                                     archivo=new_file)
                                detalleevidencia.save(request)

                            # Participantes: Autores/Coautores
                            participantes = ParticipanteSolicitudPublicacion.objects.filter(solicitud=solicitud, status=True).order_by('id')
                            for p in participantes:
                                participantepublicacion = ParticipanteCapituloLibros(matrizevidencia_id=7,
                                                                          capitulolibros=capitulo,
                                                                          profesor=p.profesor,
                                                                          tipoparticipante=p.tipo,
                                                                          administrativo=p.administrativo,
                                                                          tipoparticipanteins=p.tipoparticipanteins,
                                                                          inscripcion=p.inscripcion)
                                participantepublicacion.save(request)

                            notificar_revision_solicitud_produccion_cientifica(solicitud)

                        log(u'%s adicionó capitulo de libro de investigacion: %s' % (persona, capitulo), request, "add")
                        return JsonResponse({"result": "ok", "titulo": "Proceso exitoso!!!", "mensaje": u"Registro guardado con éxito", "showSwal": True})
                    else:
                        return JsonResponse({"result": "bad", "titulo": "Atención!!!", "mensaje": "El capítulo de libro ya ha sido ingresado", "showSwal": "True", "swalType": "warning"})
                else:
                    msg = ",".join([k + ': ' + v[0] for k, v in f.errors.items()])
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "titulo": "Error", "mensaje": u"Error al guardar los datos. [%s]" % msg, "showSwal": "True", "swalType": "error"})
            except Exception as ex:
                msg = ex.__str__()
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "titulo": "Error", "mensaje": u"Error al guardar los datos. [%s]" % msg, "showSwal": "True", "swalType": "error"})

        elif action == 'editarlibro':
            try:
                f = LibroInvestigacionForm(request.POST)

                if 'archivoderecho' in request.FILES:
                    newfile = request.FILES['archivoderecho']
                    extension = newfile._name.split('.')
                    tam = len(extension)
                    exte = extension[tam - 1]
                    if newfile.size > 10485760:
                        return JsonResponse({"result": "bad", "mensaje": u"Error, Tamaño de archivo Documento Derechos autor Máximo permitido es de 10 Mb"})
                    if exte.lower() not in ['pdf']:
                        return JsonResponse({"result": "bad", "mensaje": u"Solo se permiten archivos .pdf (Documento Derechos autor)"})

                    archivoderecho = request.FILES['archivoderecho']
                    archivoderecho._name = generar_nombre("derechoautor_", archivoderecho._name)

                libros = LibroInvestigacion.objects.get(pk=int(encrypt(request.POST['id'])))
                if f.is_valid():
                    libros.nombrelibro = f.cleaned_data['nombrelibro']
                    libros.codisbn = f.cleaned_data['codisbn']
                    libros.editorial = f.cleaned_data['editorial']
                    libros.numeroedicion = f.cleaned_data['numeroedicion']
                    libros.numeropagina = f.cleaned_data['numeropagina']
                    libros.registroderecho = f.cleaned_data['registroderecho']
                    libros.fechapublicacion = f.cleaned_data['fechapublicacion']
                    libros.revi_pare = f.cleaned_data['revi_pare']
                    libros.revi_fili = f.cleaned_data['revi_fili']
                    libros.areaconocimiento = f.cleaned_data['areaconocimiento']
                    libros.subareaconocimiento = f.cleaned_data['subareaconocimiento']
                    libros.subareaespecificaconocimiento = f.cleaned_data['subareaespecificaconocimiento']
                    libros.lineainvestigacion = f.cleaned_data['lineainvestigacion']
                    libros.sublineainvestigacion = f.cleaned_data['sublineainvestigacion']
                    libros.provieneproyecto = f.cleaned_data['provieneproyecto']
                    libros.tipoproyecto = f.cleaned_data['tipoproyecto'] if f.cleaned_data['provieneproyecto'] else None
                    libros.pertenecegrupoinv = f.cleaned_data['pertenecegrupoinv']
                    libros.grupoinvestigacion = f.cleaned_data['grupoinvestigacion']
                    libros.accesoabierto = f.cleaned_data['accesoabierto']
                    libros.save(request)

                    if f.cleaned_data['provieneproyecto']:
                        if int(f.cleaned_data['tipoproyecto']) != 3:
                            libros.proyectointerno = f.cleaned_data['proyectointerno']
                            libros.proyectoexterno = None
                        else:
                            libros.proyectoexterno = f.cleaned_data['proyectoexterno']
                            libros.proyectointerno = None
                    else:
                        libros.tipoproyecto = None
                        libros.proyectointerno = None
                        libros.proyectoexterno = None

                    libros.save(request)

                    if 'archivoderecho' in request.FILES:
                        evidencia = Evidencia.objects.get(pk=16)
                        if DetalleEvidencias.objects.filter(libro=libros, evidencia=evidencia).exists():
                            detalleevidencia = DetalleEvidencias.objects.get(libro=libros, evidencia=evidencia)
                            detalleevidencia.archivo = archivoderecho
                            detalleevidencia.save(request)
                        else:
                            detalleevidencia = DetalleEvidencias(evidencia=evidencia,
                                                                 libro=libros,
                                                                 descripcion='REGISTRO DERECHOS DE AUTOR',
                                                                 archivo=archivoderecho)
                            detalleevidencia.save(request)

                    log(u'Editó libro de investigacion: %s' % libros, request, "edit")
                    return JsonResponse({"result": "ok", "titulo": "Proceso exitoso!!!", "mensaje": u"Registro actualizado con éxito", "showSwal": True})
                else:
                     raise NameError('Error')
            except Exception as ex:
                msg = ex.__str__()
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "titulo": "Error", "mensaje": u"Error al guardar los datos. [%s]" % msg, "showSwal": "True", "swalType": "error"})

        elif action == 'editarcapitulo':
            try:
                f = CapituloLibroInvestigacionForm(request.POST)

                if f.is_valid():
                    if CapituloLibroInvestigacion.objects.values('id').filter(titulocapitulo=f.cleaned_data['titulocapitulo'], status=True).exclude(pk=int(encrypt(request.POST['id']))).exists():
                        return JsonResponse({"result": "bad", "titulo": "Atención!!!", "mensaje": "El capítulo de libro ya ha sido ingresado", "showSwal": "True", "swalType": "warning"})

                    capitulos = CapituloLibroInvestigacion.objects.get(pk=int(encrypt(request.POST['id'])))

                    capitulos.titulocapitulo = f.cleaned_data['titulocapitulo']
                    capitulos.titulolibro = f.cleaned_data['titulolibro']
                    capitulos.codisbn = f.cleaned_data['codisbn']
                    capitulos.paginas = f.cleaned_data['paginas']
                    capitulos.editorcompilador = f.cleaned_data['editorcompilador']
                    capitulos.totalcapitulo = f.cleaned_data['totalcapitulo']
                    capitulos.fechapublicacion = f.cleaned_data['fechapublicacion']
                    capitulos.filiacion = f.cleaned_data['filiacion']
                    capitulos.areaconocimiento = f.cleaned_data['areaconocimiento']
                    capitulos.subareaconocimiento = f.cleaned_data['subareaconocimiento']
                    capitulos.subareaespecificaconocimiento = f.cleaned_data['subareaespecificaconocimiento']
                    capitulos.lineainvestigacion = f.cleaned_data['lineainvestigacion']
                    capitulos.sublineainvestigacion = f.cleaned_data['sublineainvestigacion']
                    capitulos.provieneproyecto = f.cleaned_data['provieneproyecto']
                    capitulos.tipoproyecto = f.cleaned_data['tipoproyecto'] if f.cleaned_data['provieneproyecto'] else None
                    capitulos.pertenecegrupoinv = f.cleaned_data['pertenecegrupoinv']
                    capitulos.grupoinvestigacion = f.cleaned_data['grupoinvestigacion']
                    capitulos.revisadopar = f.cleaned_data['revisadopar']
                    capitulos.save(request)

                    if f.cleaned_data['provieneproyecto']:
                        if int(f.cleaned_data['tipoproyecto']) != 3:
                            capitulos.proyectointerno = f.cleaned_data['proyectointerno']
                            capitulos.proyectoexterno = None
                        else:
                            capitulos.proyectoexterno = f.cleaned_data['proyectoexterno']
                            capitulos.proyectointerno = None
                    else:
                        capitulos.tipoproyecto = None
                        capitulos.proyectointerno = None
                        capitulos.proyectoexterno = None

                    capitulos.save(request)

                    log(u'%s editó capitulo de libro de investigacion: %s' % (persona, capitulos), request, "edit")
                    return JsonResponse({"result": "ok", "titulo": "Proceso exitoso!!!", "mensaje": u"Registro actualizado con éxito", "showSwal": True})
                else:
                    msg = ",".join([k + ': ' + v[0] for k, v in f.errors.items()])
                    return JsonResponse({"result": "bad", "titulo": "Error", "mensaje": u"Error al guardar los datos. [%s]" % msg, "showSwal": "True", "swalType": "error"})
            except Exception as ex:
                msg = ex.__str__()
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "titulo": "Error", "mensaje": u"Error al guardar los datos. [%s]" % msg, "showSwal": "True", "swalType": "error"})

        elif action == 'deleteparticipantelibro':
            try:
                participante = ParticipanteLibros.objects.get(pk=request.POST['id'])
                participante.status = False
                participante.save(request)
                log(u'Eliminó participante de libro de investigacion: %s' % participante, request, "del")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        elif action == 'deleteparticipantecapitulo':
            try:
                participante = ParticipanteCapituloLibros.objects.get(pk=request.POST['id'])
                participante.status = False
                participante.save(request)
                log(u'Eliminó participante de capitulo de libro de investigacion: %s' % participante, request, "del")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        elif action == 'deletelibro':
            try:
                libros = LibroInvestigacion.objects.get(pk=request.POST['id'])
                libros.status = False
                libros.save(request)
                log(u'Eliminó libro de investigacion: %s' % libros, request, "del")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        elif action == 'deletecapitulo':
            try:
                capitulos=CapituloLibroInvestigacion.objects.get(pk=request.POST['id'])
                capitulos.status = False
                capitulos.save(request)
                log(u'Eliminó capitulo de libro de investigacion: %s' % capitulos, request, "del")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        elif action == 'addparticipantesdocentes':
            try:
                f = ParticipanteProfesorLibroForm(request.POST)
                if f.is_valid():
                    programas = ParticipanteLibros(matrizevidencia_id=6,
                                                   libros_id=int(encrypt(request.POST['id'])),
                                                   profesor_id=f.cleaned_data['profesor'],
                                                   tipoparticipante=f.cleaned_data['tipoparticipante'],
                                                   tipoparticipanteins = f.cleaned_data['tipoparticipanteins']
                                                   )
                    programas.save(request)
                    log(u'Adicionó participante de libro de investigacion: %s' % programas, request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                     raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'addparticipantescapitulodocentes':
            try:
                if request.POST['profesor'] == '0':
                    return JsonResponse({"result": "bad", "mensaje": u"Seleccione el profesor."})

                f = ParticipanteProfesorCapituloLibroForm(request.POST)

                if f.is_valid():
                    participantecapitulo = ParticipanteCapituloLibros(matrizevidencia_id=7,
                                                           capitulolibros_id=int(encrypt(request.POST['id'])),
                                                           profesor_id=f.cleaned_data['profesor'],
                                                           tipoparticipante=f.cleaned_data['tipoparticipante'],
                                                           tipoparticipanteins=f.cleaned_data['tipoparticipanteins']
                                                           )
                    participantecapitulo.save(request)
                    log(u'Adicionó participante profesor de capítulo de libro de investigacion: %s' % participantecapitulo, request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                     raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'addparticipantesadministrativos':
            try:
                f = ParticipanteAdministrativoLibroForm(request.POST)
                if f.is_valid():
                    programas = ParticipanteLibros(matrizevidencia_id=6,
                                                   libros_id=int(encrypt(request.POST['id'])),
                                                   administrativo_id=f.cleaned_data['administrativo'],
                                                   tipoparticipante=f.cleaned_data['tipoparticipante'],
                                                   tipoparticipanteins=f.cleaned_data['tipoparticipanteins']
                                                   )
                    programas.save(request)
                    log(u'Adicionó participante administrativo de libro de investigacion: %s' % programas, request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                     raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'addparticipantescapituloadministrativos':
            try:
                if request.POST['administrativo'] == '0':
                    return JsonResponse({"result": "bad", "mensaje": u"Seleccione un administrativo."})

                f = ParticipanteAdministrativoCapituloLibroForm(request.POST)
                if f.is_valid():
                    participantecapitulo = ParticipanteCapituloLibros(matrizevidencia_id=7,
                                                           capitulolibros_id=int(encrypt(request.POST['id'])),
                                                           administrativo_id=f.cleaned_data['administrativo'],
                                                           tipoparticipante=f.cleaned_data['tipoparticipante'],
                                                           tipoparticipanteins=f.cleaned_data['tipoparticipanteins']
                                                           )
                    participantecapitulo.save(request)
                    log(u'Adicionó participante administrativo a capitulo de libro de investigacion: %s' % participantecapitulo, request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                     raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'addevidenciaslibro':
            try:
                f = EvidenciaForm(request.POST, request.FILES)
                # 2.5MB - 2621440
                # 5MB - 5242880
                # 10MB - 10485760
                # 20MB - 20971520
                # 50MB - 5242880
                # 100MB 104857600
                # 250MB - 214958080
                # 500MB - 429916160
                d = request.FILES['archivo']
                if d.size > 10485760:
                    return JsonResponse({"result": "bad", "mensaje": u"Error, archivo mayor a 10 Mb."})
                if f.is_valid():
                    newfile = None
                    if 'archivo' in request.FILES:
                        newfile = request.FILES['archivo']
                        newfile._name = generar_nombre("libro_", newfile._name)
                    if DetalleEvidencias.objects.values('id').filter(evidencia_id=int(encrypt(request.POST['idevidencia'])),
                                                        libro_id=int(encrypt(request.POST['id']))).exists():
                        detalle = DetalleEvidencias.objects.get(evidencia_id=int(encrypt(request.POST['idevidencia'])),
                                                                libro_id=int(encrypt(request.POST['id'])))
                        detalle.descripcion = f.cleaned_data['descripcion']
                        detalle.archivo = newfile
                        detalle.save(request)
                        log(u'Adicionó evidencias de libro de investigacion: %s' % detalle, request, "add")
                    else:
                        evidencia = DetalleEvidencias(evidencia_id=int(encrypt(request.POST['idevidencia'])),
                                                      libro_id=int(encrypt(request.POST['id'])),
                                                      descripcion=f.cleaned_data['descripcion'],
                                                      archivo=newfile)
                        evidencia.save(request)
                        log(u'Adicionó evidencias de libro de investigacion: %s' % evidencia, request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                     raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'addevidenciascapitulo':
            try:
                f = EvidenciaForm(request.POST, request.FILES)
                # 2.5MB - 2621440
                # 5MB - 5242880
                # 10MB - 10485760
                # 20MB - 20971520
                # 50MB - 5242880
                # 100MB 104857600
                # 250MB - 214958080
                # 500MB - 429916160
                d = request.FILES['archivo']
                if d.size > 10485760:
                    return JsonResponse({"result": "bad", "mensaje": u"Error, archivo mayor a 10 Mb."})
                if f.is_valid():
                    newfile = None
                    if 'archivo' in request.FILES:
                        newfile = request.FILES['archivo']
                        newfile._name = generar_nombre("evidenciacapitulo_", newfile._name)
                    if DetalleEvidencias.objects.values('id').filter(evidencia_id=int(encrypt(request.POST['idevidencia'])),
                                                        capitulo_id=int(encrypt(request.POST['id']))).exists():
                        detalle = DetalleEvidencias.objects.get(evidencia_id=int(encrypt(request.POST['idevidencia'])),
                                                                capitulo_id=int(encrypt(request.POST['id'])))
                        detalle.descripcion = f.cleaned_data['descripcion']
                        detalle.archivo = newfile
                        detalle.save(request)
                        log(u'Adicionó evidencias de capitulo de libro de investigacion: %s' % detalle, request, "add")
                    else:
                        evidencia = DetalleEvidencias(evidencia_id=int(encrypt(request.POST['idevidencia'])),
                                                      capitulo_id=int(encrypt(request.POST['id'])),
                                                      descripcion=f.cleaned_data['descripcion'],
                                                      archivo=newfile)
                        evidencia.save(request)
                        log(u'Adicionó evidencias de capitulo de libro de investigacion: %s' % evidencia, request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                     raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'solicitudes':
            try:
                data['solicitudes'] = SolicitudPublicacion.objects.filter(tiposolicitud=int(request.POST['tipo']), status=True, estado__valor=1).order_by('fecha_creacion')

                template = get_template("inv_libros/solicitudes.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'html': json_content})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'aprobar':
            try:
                solicitudes = SolicitudPublicacion.objects.get(pk=request.POST['id'])
                solicitudes.aprobado = True
                solicitudes.save(request)
                log(u'Aprobo solicitud: %s' % solicitudes.persona, request, "edit")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                pass
                return JsonResponse({"result": "bad", "mensaje": u"Error al aprobar los datos."})

        elif action == 'updatetipoparticipante':
            try:
                participantes = ParticipanteLibros.objects.get(pk=request.POST['iditem'])
                participantes.tipoparticipanteins = request.POST['idtipo']
                participantes.save(request)
                return JsonResponse({'result': 'ok', 'id': participantes.libros.id})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad'})

        elif action == 'aprobarlibro':
            try:
                libro = LibroInvestigacion.objects.get(pk=int(encrypt(request.POST['id'])))

                # Si existe solicitud
                if libro.solicitudpublicacion:
                    estadosolicitud = obtener_estado_solicitud(8, 5)  # APROBADO

                    solicitud = SolicitudPublicacion.objects.get(pk=libro.solicitudpublicacion.id)
                    solicitud.aprobado = True
                    solicitud.estado = estadosolicitud
                    solicitud.save(request)

                libro.aprobado = True
                libro.save(request)

                integrantes = libro.participantes()
                libro.tipoaporte = 1 if integrantes.filter(tipoparticipanteins=1).count() >= 1 else 2
                libro.save(request)

                # Si existe solicitud se debe notificar
                if libro.solicitudpublicacion:
                    notificar_revision_solicitud_produccion_cientifica(solicitud)

                log(u'%s Aprobó libro: %s' % (persona, libro.nombrelibro), request, "edit")
                return JsonResponse({"result": "ok", "titulo": "Proceso exitoso!!!", "mensaje": u"Registro guardado con éxito", "showSwal": True})
                # return JsonResponse({"result": "ok"})
            except Exception as ex:
                msg = ex.__str__()
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "titulo": "Error", "mensaje": u"Error al guardar los datos. [%s]" % msg, "showSwal": "True", "swalType": "error"})

        elif action == 'aprobarcapitulolibro':
            try:
                capitulo = CapituloLibroInvestigacion.objects.get(pk=int(encrypt(request.POST['id'])))

                # Si existe solicitud
                if capitulo.solicitudpublicacion:
                    estadosolicitud = obtener_estado_solicitud(8, 5)  # APROBADO

                    solicitud = SolicitudPublicacion.objects.get(pk=capitulo.solicitudpublicacion.id)
                    solicitud.aprobado = True
                    solicitud.estado = estadosolicitud
                    solicitud.save(request)

                capitulo.aprobado = True
                capitulo.save(request)

                integrantes = capitulo.participantes()
                capitulo.tipoaporte = 1 if integrantes.filter(tipoparticipanteins=1).count() >= 1 else 2
                capitulo.save(request)

                # Si existe solicitud se debe notificar
                if capitulo.solicitudpublicacion:
                    notificar_revision_solicitud_produccion_cientifica(solicitud)

                log(u'%s Aprobó capítulo de libro: %s' % (persona, capitulo.titulocapitulo), request, "edit")
                return JsonResponse({"result": "ok", "titulo": "Proceso exitoso!!!", "mensaje": u"Registro guardado con éxito", "showSwal": True})
            except Exception as ex:
                msg = ex.__str__()
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "titulo": "Error", "mensaje": u"Error al guardar los datos. [%s]" % msg, "showSwal": "True", "swalType": "error"})

        elif action == 'rechazar':
            try:
                solicitudes = SolicitudPublicacion.objects.get(pk=request.POST['id'])
                observacion=request.POST['observacion']
                if observacion:
                    solicitudes.aprobado = False
                    solicitudes.observacion=observacion
                    solicitudes.save(request)
                    log(u'Rechazo solicitud: %s' % solicitudes.persona, request, "edit")
                    asunto = u"ESTADO DE APROBACIÓN DE LIBRO"
                    send_html_mail(asunto, "emails/estadosolicitudarticulo.html", {'sistema': request.session['nombresistema'],'solicitudes':solicitudes,'profesor': solicitudes.persona.nombre_completo_inverso()}, solicitudes.persona.lista_emails_envio(), [], cuenta=CUENTAS_CORREOS[4][1])
                    return JsonResponse({"result": "ok"})
                else:
                    return JsonResponse({"result": "bad", "mensaje": u"Debe ingresar observación."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al aprobar los datos."})

        elif action == 'reportefichacatalografica':
            try:
                data = {}
                libro = LibroInvestigacion.objects.get(pk=int(encrypt(request.POST['id'])))
                participantes = libro.participantes()
                data['libro'] = libro
                data['participantes'] = participantes

                return conviert_html_to_pdf(
                    'inv_libros/fichacatalografica_pdf.html',
                    {
                        'pagesize': 'A4',
                        'data': data,
                    }
                )
            except Exception as ex:
                return HttpResponseRedirect("/ponenciasinvestigacion?info=%s" % "Error al generar el reporte de ficha catalográfica")

        elif action == 'reportefichacatalograficacapitulo':
            try:
                data = {}
                capitulo = CapituloLibroInvestigacion.objects.get(pk=int(encrypt(request.POST['id'])))
                data['capitulo'] = capitulo
                data['participantes'] = capitulo.participantes()

                return conviert_html_to_pdf(
                    'inv_libros/fichacatalograficacapitulo_pdf.html',
                    {
                        'pagesize': 'A4',
                        'data': data,
                    }
                )
            except Exception as ex:
                return HttpResponseRedirect("/librosinvestigacion?info=%s" % "Error al generar el reporte de ficha catalográfica")

        elif action == 'updatetipoparticipantecapitulo':
            try:
                participantes = ParticipanteCapituloLibros.objects.get(pk=request.POST['iditem'])
                participantes.tipoparticipanteins = request.POST['idtipo']
                participantes.save(request)
                return JsonResponse({'result': 'ok', 'id': participantes.capitulolibros.id})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad'})

        elif action == 'evidenciaslibro':
            try:
                # Aquí se almacenan las fichas catalográficas
                directorio = SITE_STORAGE + '/media/articulos'
                try:
                    os.stat(directorio)
                except:
                    os.mkdir(directorio)

                caracteres_a_quitar = "!\"\#$%&()=?¡'¿{}[],.-+*/|°^~:;"
                anio = request.POST['anio']
                output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media', 'zipav'))

                response = HttpResponse(content_type='application/zip')
                response['Content-Disposition'] = 'attachment; filename=librosevidencias' + random.randint(1, 10000).__str__() + '.zip'

                nombre_archivo = "evidenciaslibros" + anio + ".zip"
                filename = os.path.join(output_folder, nombre_archivo)
                fantasy_zip = zipfile.ZipFile(filename, 'w')

                libros = LibroInvestigacion.objects.filter(status=True, fechapublicacion__year=anio).order_by('id')

                for libro in libros:
                    titulo = libro.nombrelibro

                    palabras = titulo.split(" ")
                    titulo = "_".join(palabras[0:5])
                    titulo = remover_caracteres(titulo, caracteres_a_quitar)

                    titulo = elimina_tildes(remover_caracteres_tildes_unicode(remover_caracteres_especiales_unicode(titulo)))
                    carpetalibro = "LIB_" + libro.codisbn.replace("-", "_") + "_"+ str(libro.id) + "_" + titulo

                    libro_id = libro.id
                    nombre = "LIBRO_" + str(libro_id).zfill(4)

                    # Agrego las evidencias a la carpeta del libro
                    for evidencia in libro.detalleevidencias_set.filter(status=True):
                        ext = evidencia.archivo.__str__()[evidencia.archivo.__str__().rfind("."):]

                        if evidencia.descripcion:
                            nombreevidencia = evidencia.descripcion.upper()
                            nombreevidencia = remover_caracteres(nombreevidencia, caracteres_a_quitar)
                            nombreevidencia = elimina_tildes(remover_caracteres_tildes_unicode(remover_caracteres_especiales_unicode(nombreevidencia))).replace(" ", "_")
                        else:
                            nombreevidencia = "EVIDENCIA" + random.randint(1, 10000).__str__()


                        if evidencia.evidencia.id == 7:
                            if os.path.exists(SITE_STORAGE + evidencia.archivo.url):
                                fantasy_zip.write(SITE_STORAGE + evidencia.archivo.url, carpetalibro + '/' + nombreevidencia + ext.lower())
                        elif evidencia.evidencia.id == 16:
                            if os.path.exists(SITE_STORAGE + evidencia.archivo.url):
                                fantasy_zip.write(SITE_STORAGE + evidencia.archivo.url, carpetalibro + '/' + nombreevidencia + ext.lower())

                    # Generar la ficha catalográfica del libro
                    data['libro'] = libro
                    data['participantes'] = libro.participantes()

                    nombrearchivoficha = 'fichacatalograficalib_' + str(libro_id) + '_' + str(random.randint(1, 10000)) + '.pdf'
                    valida = convert_html_to_pdf(
                        'inv_libros/fichacatalografica_pdf.html',
                        {'pagesize': 'A4', 'data': data},
                        nombrearchivoficha,
                        directorio
                    )

                    archivoficha = directorio + "/" + nombrearchivoficha
                    # Agrego el archivo de la ficha a la carpeta del artículo
                    fantasy_zip.write(archivoficha, carpetalibro + "/FICHA_CATALOGRAFICA.pdf")
                    # Borro el archivo de la ficha creado
                    os.remove(archivoficha)

                fantasy_zip.close()

                ruta = "media/zipav/" + nombre_archivo

                return JsonResponse({'result': 'ok', 'archivo': ruta})
            except Exception as ex:
                msg = ex.__str__()
                return JsonResponse({"result": "bad", "mensaje": u"Error al descargar evidencias. Detalle: %s" % (msg)})

        elif action == 'evidenciaslibrocodigo':
            try:
                archivo = request.FILES['archivo']
                descripcionarchivo = 'Archivo Excel de Códigos'
                resp = validar_archivo(descripcionarchivo, archivo, ['xlsx'], '4MB')
                if resp['estado'] != "OK":
                    return JsonResponse({"result": "bad", "titulo": "Atención!!!", "mensaje": resp["mensaje"], "showSwal": "True", "swalType": "warning"})

                # Aquí se almacenan las fichas catalográficas
                directorio = SITE_STORAGE + '/media/articulos'
                try:
                    os.stat(directorio)
                except:
                    os.mkdir(directorio)

                caracteres_a_quitar = "!\"\#$%&()=?¡'¿{}[],.-+*/|°^~:;"
                output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media', 'zipav'))

                response = HttpResponse(content_type='application/zip')
                response['Content-Disposition'] = 'attachment; filename=librosevidencias' + random.randint(1, 10000).__str__() + '.zip'

                nombre_archivo = "evidenciaslibroscodigo.zip"
                filename = os.path.join(output_folder, nombre_archivo)
                fantasy_zip = zipfile.ZipFile(filename, 'w')

                wb = openpyxl.load_workbook(archivo)
                sheet = wb.worksheets[0]
                lista_ids = []
                lista_filas = []
                novalidos = 0

                # Recorrer las filas de la hoja
                for fila in range(1, sheet.max_row + 1):
                    if fila > 1:
                        codigo = sheet.cell(row=fila, column=1).value
                        if codigo:
                            codigo = codigo.strip()
                            if codigo_publicacion_valido(codigo, "LIB"):
                                lista_ids.append(codigo.split(" ")[1].split("-")[0])
                            else:
                                lista_filas.append(fila)
                                novalidos += 1

                if novalidos > 0:
                    return JsonResponse({"result": "bad", "titulo": "Atención!!!", "mensaje": "{} {} {} que {} el código en formato incorrecto. Filas: {}".format("Existe" if novalidos == 1 else "Existen", novalidos, "registro" if novalidos == 1 else "registros", "tiene" if novalidos == 1 else "tienen", ", ".join(str(f) for f in lista_filas)), "showSwal": "True", "swalType": "warning"})

                anios = LibroInvestigacion.objects.filter(Q(status=True) | Q(status=False, eliminadoxdoc=True), fechapublicacion__isnull=False).annotate(anio=ExtractYear('fechapublicacion')).values_list('anio', flat=True).order_by('-anio').distinct()

                for anio in anios:
                    libros = LibroInvestigacion.objects.filter(status=True, fechapublicacion__year=anio, pk__in=lista_ids).order_by('id')
                    if libros:
                        for libro in libros:
                            titulo = libro.nombrelibro

                            palabras = titulo.split(" ")
                            titulo = "_".join(palabras[0:5])
                            titulo = remover_caracteres(titulo, caracteres_a_quitar)

                            titulo = elimina_tildes(remover_caracteres_tildes_unicode(remover_caracteres_especiales_unicode(titulo)))
                            carpetalibro = "ANIO_" + str(anio) + "/LIB_" + libro.codisbn.replace("-", "_") + "_"+ str(libro.id) + "_" + titulo

                            libro_id = libro.id
                            nombre = "LIBRO_" + str(libro_id).zfill(4)

                            # Agrego las evidencias a la carpeta del libro
                            for evidencia in libro.detalleevidencias_set.filter(status=True):
                                ext = evidencia.archivo.__str__()[evidencia.archivo.__str__().rfind("."):]

                                if evidencia.descripcion:
                                    nombreevidencia = evidencia.descripcion.upper()
                                    nombreevidencia = remover_caracteres(nombreevidencia, caracteres_a_quitar)
                                    nombreevidencia = elimina_tildes(remover_caracteres_tildes_unicode(remover_caracteres_especiales_unicode(nombreevidencia))).replace(" ", "_")
                                else:
                                    nombreevidencia = "EVIDENCIA" + random.randint(1, 10000).__str__()


                                if evidencia.evidencia.id == 7:
                                    if os.path.exists(SITE_STORAGE + evidencia.archivo.url):
                                        fantasy_zip.write(SITE_STORAGE + evidencia.archivo.url, carpetalibro + '/' + nombreevidencia + ext.lower())
                                elif evidencia.evidencia.id == 16:
                                    if os.path.exists(SITE_STORAGE + evidencia.archivo.url):
                                        fantasy_zip.write(SITE_STORAGE + evidencia.archivo.url, carpetalibro + '/' + nombreevidencia + ext.lower())

                            # Generar la ficha catalográfica del libro
                            data['libro'] = libro
                            data['participantes'] = libro.participantes()

                            nombrearchivoficha = 'fichacatalograficalib_' + str(libro_id) + '_' + str(random.randint(1, 10000)) + '.pdf'
                            valida = convert_html_to_pdf(
                                'inv_libros/fichacatalografica_pdf.html',
                                {'pagesize': 'A4', 'data': data},
                                nombrearchivoficha,
                                directorio
                            )

                            archivoficha = directorio + "/" + nombrearchivoficha
                            # Agrego el archivo de la ficha a la carpeta del artículo
                            fantasy_zip.write(archivoficha, carpetalibro + "/FICHA_CATALOGRAFICA.pdf")
                            # Borro el archivo de la ficha creado
                            os.remove(archivoficha)

                fantasy_zip.close()
                ruta = "media/zipav/" + nombre_archivo
                return JsonResponse({'result': 'ok', 'archivo': ruta})
            except Exception as ex:
                msg = ex.__str__()
                return JsonResponse({"result": "bad", "titulo": "Error", "mensaje": u"Error al descargar evidencias. Detalle: [%s]" % msg, "showSwal": "True", "swalType": "error"})

        elif action == 'evidenciascapitulolibro':
            try:
                # Aquí se almacenan las fichas catalográficas
                directorio = SITE_STORAGE + '/media/articulos'
                try:
                    os.stat(directorio)
                except:
                    os.mkdir(directorio)

                caracteres_a_quitar = "!\"\#$%&()=?¡'¿{}[],.-+*/|°^~:;"
                anio = request.POST['anio']
                output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media', 'zipav'))

                response = HttpResponse(content_type='application/zip')
                response['Content-Disposition'] = 'attachment; filename=capituloslibrosevidencias' + random.randint(1, 10000).__str__() + '.zip'

                nombre_archivo = "evidenciascapituloslibros" + anio + ".zip"
                filename = os.path.join(output_folder, nombre_archivo)
                fantasy_zip = zipfile.ZipFile(filename, 'w')

                capitulos = CapituloLibroInvestigacion.objects.filter(status=True, fechapublicacion__year=anio).order_by('id')

                for capitulo in capitulos:
                    titulo = capitulo.titulocapitulo

                    palabras = titulo.split(" ")
                    titulo = "_".join(palabras[0:5])
                    titulo = remover_caracteres(titulo, caracteres_a_quitar)

                    titulo = elimina_tildes(remover_caracteres_tildes_unicode(remover_caracteres_especiales_unicode(titulo)))
                    carpetacapitulo = "CAPLIB_" + capitulo.codisbn.replace("-", "_") + "_"+ str(capitulo.id) + "_" + titulo

                    capitulo_id = capitulo.id
                    nombre = "CAPITULO_LIBRO_" + str(capitulo_id).zfill(4)

                    # Agrego las evidencias a la carpeta del capítulo de libro
                    for evidencia in capitulo.detalleevidencias_set.filter(status=True):
                        ext = evidencia.archivo.__str__()[evidencia.archivo.__str__().rfind("."):]

                        if evidencia.descripcion:
                            nombreevidencia = evidencia.descripcion.upper()
                            nombreevidencia = remover_caracteres(nombreevidencia, caracteres_a_quitar)
                            nombreevidencia = elimina_tildes(remover_caracteres_tildes_unicode(remover_caracteres_especiales_unicode(nombreevidencia))).replace(" ", "_")
                        else:
                            nombreevidencia = "EVIDENCIA" + random.randint(1, 10000).__str__()


                        if evidencia.evidencia.id == 11:
                            if os.path.exists(SITE_STORAGE + evidencia.archivo.url):
                                fantasy_zip.write(SITE_STORAGE + evidencia.archivo.url, carpetacapitulo + '/' + nombreevidencia + ext.lower())

                        elif evidencia.evidencia.id == 12:
                            if os.path.exists(SITE_STORAGE + evidencia.archivo.url):
                                fantasy_zip.write(SITE_STORAGE + evidencia.archivo.url, carpetacapitulo + '/' + nombreevidencia + ext.lower())

                        elif evidencia.evidencia.id == 14:
                            if os.path.exists(SITE_STORAGE + evidencia.archivo.url):
                                fantasy_zip.write(SITE_STORAGE + evidencia.archivo.url, carpetacapitulo + '/' + nombreevidencia + ext.lower())

                        elif evidencia.evidencia.id == 13:
                            if os.path.exists(SITE_STORAGE + evidencia.archivo.url):
                                fantasy_zip.write(SITE_STORAGE + evidencia.archivo.url, carpetacapitulo + '/' + nombreevidencia + ext.lower())

                    # Generar la ficha catalográfica del capítulo de libro
                    data['capitulo'] = capitulo
                    data['participantes'] = capitulo.participantes()

                    nombrearchivoficha = 'fichacatalograficacap_' + str(capitulo_id) + '_' + str(random.randint(1, 10000)) + '.pdf'
                    valida = convert_html_to_pdf(
                        'inv_libros/fichacatalograficacapitulo_pdf.html',
                        {'pagesize': 'A4', 'data': data},
                        nombrearchivoficha,
                        directorio
                    )

                    archivoficha = directorio + "/" + nombrearchivoficha
                    # Agrego el archivo de la ficha a la carpeta del artículo
                    fantasy_zip.write(archivoficha, carpetacapitulo + "/FICHA_CATALOGRAFICA.pdf")
                    # Borro el archivo de la ficha creado
                    os.remove(archivoficha)


                fantasy_zip.close()

                ruta = "media/zipav/" + nombre_archivo

                return JsonResponse({'result': 'ok', 'archivo': ruta})
            except Exception as ex:
                msg = ex.__str__()
                return JsonResponse({"result": "bad", "mensaje": u"Error al descargar evidencias. Detalle: %s" % (msg)})

        elif action == 'evidenciascapitulocodigo':
            try:
                archivo = request.FILES['archivo']
                descripcionarchivo = 'Archivo Excel de Códigos'
                resp = validar_archivo(descripcionarchivo, archivo, ['xlsx'], '4MB')
                if resp['estado'] != "OK":
                    return JsonResponse({"result": "bad", "titulo": "Atención!!!", "mensaje": resp["mensaje"], "showSwal": "True", "swalType": "warning"})

                # Aquí se almacenan las fichas catalográficas
                directorio = SITE_STORAGE + '/media/articulos'
                try:
                    os.stat(directorio)
                except:
                    os.mkdir(directorio)

                caracteres_a_quitar = "!\"\#$%&()=?¡'¿{}[],.-+*/|°^~:;"
                output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media', 'zipav'))

                response = HttpResponse(content_type='application/zip')
                response['Content-Disposition'] = 'attachment; filename=capituloslibrosevidencias' + random.randint(1, 10000).__str__() + '.zip'

                nombre_archivo = "evidenciascapituloslibroscodigo.zip"
                filename = os.path.join(output_folder, nombre_archivo)
                fantasy_zip = zipfile.ZipFile(filename, 'w')

                wb = openpyxl.load_workbook(archivo)
                sheet = wb.worksheets[0]
                lista_ids = []
                lista_filas = []
                novalidos = 0
                # Recorrer las filas de la hoja
                for fila in range(1, sheet.max_row + 1):
                    if fila > 1:
                        codigo = sheet.cell(row=fila, column=1).value
                        if codigo:
                            codigo = codigo.strip()
                            if codigo_publicacion_valido(codigo, "CAP"):
                                lista_ids.append(codigo.split(" ")[1].split("-")[0])
                            else:
                                lista_filas.append(fila)
                                novalidos += 1

                if novalidos > 0:
                    return JsonResponse({"result": "bad", "titulo": "Atención!!!", "mensaje": "{} {} {} que {} el código en formato incorrecto. Filas: {}".format("Existe" if novalidos == 1 else "Existen", novalidos, "registro" if novalidos == 1 else "registros", "tiene" if novalidos == 1 else "tienen", ", ".join(str(f) for f in lista_filas)), "showSwal": "True", "swalType": "warning"})

                anios = CapituloLibroInvestigacion.objects.filter(Q(status=True) | Q(status=False, eliminadoxdoc=True), fechapublicacion__isnull=False).annotate(anio=ExtractYear('fechapublicacion')).values_list('anio', flat=True).order_by('-anio').distinct()

                for anio in anios:
                    capitulos = CapituloLibroInvestigacion.objects.filter(status=True, fechapublicacion__year=anio, pk__in=lista_ids).order_by('id')

                    if capitulos:
                        for capitulo in capitulos:
                            titulo = capitulo.titulocapitulo

                            palabras = titulo.split(" ")
                            titulo = "_".join(palabras[0:5])
                            titulo = remover_caracteres(titulo, caracteres_a_quitar)

                            titulo = elimina_tildes(remover_caracteres_tildes_unicode(remover_caracteres_especiales_unicode(titulo)))
                            carpetacapitulo = "ANIO_" + str(anio) + "/CAPLIB_" + capitulo.codisbn.replace("-", "_") + "_"+ str(capitulo.id) + "_" + titulo

                            capitulo_id = capitulo.id
                            nombre = "CAPITULO_LIBRO_" + str(capitulo_id).zfill(4)

                            # Agrego las evidencias a la carpeta del capítulo de libro
                            for evidencia in capitulo.detalleevidencias_set.filter(status=True):
                                ext = evidencia.archivo.__str__()[evidencia.archivo.__str__().rfind("."):]

                                if evidencia.descripcion:
                                    nombreevidencia = evidencia.descripcion.upper()
                                    nombreevidencia = remover_caracteres(nombreevidencia, caracteres_a_quitar)
                                    nombreevidencia = elimina_tildes(remover_caracteres_tildes_unicode(remover_caracteres_especiales_unicode(nombreevidencia))).replace(" ", "_")
                                else:
                                    nombreevidencia = "EVIDENCIA" + random.randint(1, 10000).__str__()


                                if evidencia.evidencia.id == 11:
                                    if os.path.exists(SITE_STORAGE + evidencia.archivo.url):
                                        fantasy_zip.write(SITE_STORAGE + evidencia.archivo.url, carpetacapitulo + '/' + nombreevidencia + ext.lower())

                                elif evidencia.evidencia.id == 12:
                                    if os.path.exists(SITE_STORAGE + evidencia.archivo.url):
                                        fantasy_zip.write(SITE_STORAGE + evidencia.archivo.url, carpetacapitulo + '/' + nombreevidencia + ext.lower())

                                elif evidencia.evidencia.id == 14:
                                    if os.path.exists(SITE_STORAGE + evidencia.archivo.url):
                                        fantasy_zip.write(SITE_STORAGE + evidencia.archivo.url, carpetacapitulo + '/' + nombreevidencia + ext.lower())

                                elif evidencia.evidencia.id == 13:
                                    if os.path.exists(SITE_STORAGE + evidencia.archivo.url):
                                        fantasy_zip.write(SITE_STORAGE + evidencia.archivo.url, carpetacapitulo + '/' + nombreevidencia + ext.lower())

                            # Generar la ficha catalográfica del capítulo de libro
                            data['capitulo'] = capitulo
                            data['participantes'] = capitulo.participantes()

                            nombrearchivoficha = 'fichacatalograficacap_' + str(capitulo_id) + '_' + str(random.randint(1, 10000)) + '.pdf'
                            valida = convert_html_to_pdf(
                                'inv_libros/fichacatalograficacapitulo_pdf.html',
                                {'pagesize': 'A4', 'data': data},
                                nombrearchivoficha,
                                directorio
                            )

                            archivoficha = directorio + "/" + nombrearchivoficha
                            # Agrego el archivo de la ficha a la carpeta del artículo
                            fantasy_zip.write(archivoficha, carpetacapitulo + "/FICHA_CATALOGRAFICA.pdf")
                            # Borro el archivo de la ficha creado
                            os.remove(archivoficha)

                fantasy_zip.close()
                ruta = "media/zipav/" + nombre_archivo
                return JsonResponse({'result': 'ok', 'archivo': ruta})
            except Exception as ex:
                msg = ex.__str__()
                return JsonResponse({"result": "bad", "titulo": "Error", "mensaje": u"Error al descargar evidencias. Detalle: %s" % (msg)})


        return JsonResponse({"result": "bad", "mensaje": u"Solicitud Incorrecta."})
    else:
        if 'action' in request.GET:
            action = request.GET['action']

            if action == 'aprobar':
                try:
                    data['title'] = u'Aprobar Solicitud Articulos'
                    data['solicitudes'] = SolicitudPublicacion.objects.get(pk=request.GET['id'])
                    return render(request, "inv_libros/aprobar.html", data)
                except Exception as ex:
                    pass

            elif action == 'aprobarlibro':
                try:
                    data['title'] = u'Aprobar Libros'
                    libro = LibroInvestigacion.objects.get(pk=request.GET['idlibro'])
                    data['libro'] = libro
                    return render(request, "inv_libros/aprobarlibro.html", data)
                except Exception as ex:
                    pass

            elif action == 'aprobarcapitulolibro':
                try:
                    data['title'] = u'Aprobar Capítulos'
                    capitulo = CapituloLibroInvestigacion.objects.get(pk=request.GET['idcapitulo'])
                    data['capitulo'] = capitulo
                    return render(request, "inv_libros/aprobarcapitulo.html", data)
                except Exception as ex:
                    pass

            elif action == 'add':
                try:
                    data['title'] = u'Adicionar Libros'
                    form = LibroInvestigacionForm()
                    data['form'] = form
                    return render(request, "inv_libros/add.html", data)
                except Exception as ex:
                    pass

            elif action == 'ingresarsolicitudlibro':
                try:
                    data['title'] = u'Adicionar Libros'
                    solicitud = SolicitudPublicacion.objects.get(pk=int(encrypt(request.GET['id'])))
                    form = LibroInvestigacionForm(initial={
                        'nombrelibro': solicitud.nombre,
                        'codisbn': solicitud.codigoisbn,
                        'editorial': solicitud.editorcompilador,
                        'fechapublicacion': solicitud.fechapublicacion,
                        'revi_pare': 1 if solicitud.revisadopar else 2,
                        'areaconocimiento': solicitud.areaconocimiento,
                        'subareaconocimiento': solicitud.subareaconocimiento,
                        'subareaespecificaconocimiento': solicitud.subareaespecificaconocimiento,
                        'lineainvestigacion': solicitud.lineainvestigacion,
                        'sublineainvestigacion': solicitud.sublineainvestigacion,
                        'provieneproyecto': solicitud.provieneproyecto,
                        'tipoproyecto': solicitud.tipoproyecto,
                        'proyectointerno': solicitud.proyectointerno,
                        'proyectoexterno': solicitud.proyectoexterno,
                        'pertenecegrupoinv': solicitud.pertenecegrupoinv,
                        'grupoinvestigacion': solicitud.grupoinvestigacion
                    })
                    form.editar(solicitud)
                    data['form'] = form
                    data['ids'] = solicitud.id
                    data['evidencias'] = solicitud.evidencias()
                    data['participantes'] = solicitud.participantes()
                    data['estadossolicitud'] = obtener_estados_solicitud(8, [2, 3, 4])

                    return render(request, "inv_libros/ingresarsolicitudlibro.html", data)
                except Exception as ex:
                    pass

            elif action == 'addcapitulo':
                try:
                    data['title'] = u'Adicionar Capítulo de Libro'
                    if 'ids' in request.GET:
                        solicitud = SolicitudPublicacion.objects.get(pk=int(encrypt(request.GET['ids'])))
                        data['ids'] = solicitud.id
                        form = CapituloLibroInvestigacionForm(initial={
                            'titulocapitulo': solicitud.nombre if solicitud.nombre else solicitud.motivo,
                            'titulolibro': solicitud.evento,
                            'codisbn': solicitud.codigoisbn,
                            'paginas': solicitud.paginas,
                            'editorcompilador': solicitud.editorcompilador,
                            'totalcapitulo': solicitud.totalcapitulo,
                            'fechapublicacion': solicitud.fechapublicacion,
                            'filiacion': solicitud.filiacion,
                            'areaconocimiento': solicitud.areaconocimiento,
                            'subareaconocimiento': solicitud.subareaconocimiento,
                            'subareaespecificaconocimiento': solicitud.subareaespecificaconocimiento,
                            'lineainvestigacion': solicitud.lineainvestigacion,
                            'sublineainvestigacion': solicitud.sublineainvestigacion,
                            'provieneproyecto': solicitud.provieneproyecto,
                            'tipoproyecto': solicitud.tipoproyecto,
                            'proyectointerno': solicitud.proyectointerno,
                            'proyectoexterno': solicitud.proyectoexterno,
                            'pertenecegrupoinv': solicitud.pertenecegrupoinv,
                            'grupoinvestigacion': solicitud.grupoinvestigacion,
                            'revisadopar': solicitud.revisadopar
                        })
                        form.editar(solicitud)

                        data['evidencias'] = solicitud.evidencias()
                        data['participantes'] = solicitud.participantes()
                        data['estadossolicitud'] = obtener_estados_solicitud(8, [2, 3, 4])
                    else:
                        form = CapituloLibroInvestigacionForm()
                    data['form'] = form
                    return render(request, "inv_libros/addcapitulo.html", data)
                except Exception as ex:
                    pass

            elif action == 'editarlibro':
                try:
                    data['title'] = u'Editar Libro'
                    data['libros'] = libros = LibroInvestigacion.objects.get(pk=int(encrypt(request.GET['id'])))
                    form = LibroInvestigacionForm(initial={'nombrelibro': libros.nombrelibro,
                                                           'codisbn': libros.codisbn,
                                                           'editorial': libros.editorial,
                                                           'fechapublicacion': libros.fechapublicacion,
                                                           'numeroedicion': libros.numeroedicion,
                                                           'numeropagina': libros.numeropagina,
                                                           'registroderecho': libros.registroderecho,
                                                           'revi_pare': libros.revi_pare,
                                                           'revi_fili': libros.revi_fili,
                                                           'areaconocimiento': libros.areaconocimiento,
                                                           'subareaconocimiento': libros.subareaconocimiento,
                                                           'subareaespecificaconocimiento': libros.subareaespecificaconocimiento,
                                                           'lineainvestigacion': libros.lineainvestigacion,
                                                           'sublineainvestigacion': libros.sublineainvestigacion,
                                                           'provieneproyecto': libros.provieneproyecto,
                                                           'tipoproyecto': libros.tipoproyecto,
                                                           'proyectointerno': libros.proyectointerno,
                                                           'proyectoexterno': libros.proyectoexterno,
                                                           'pertenecegrupoinv': libros.pertenecegrupoinv,
                                                           'grupoinvestigacion': libros.grupoinvestigacion,
                                                           'accesoabierto': libros.accesoabierto})
                    form.editar(libros)
                    data['form'] = form
                    return render(request, "inv_libros/editarlibro.html", data)
                except Exception as ex:
                    pass

            elif action == 'editarcapitulo':
                try:
                    data['title'] = u'Editar Capítulo'
                    data['capitulo'] = capitulos = CapituloLibroInvestigacion.objects.get(pk=int(encrypt(request.GET['id'])))
                    form = CapituloLibroInvestigacionForm(initial={'titulocapitulo': capitulos.titulocapitulo,
                                                                   'titulolibro': capitulos.titulolibro,
                                                                   'codisbn': capitulos.codisbn,
                                                                   'paginas': capitulos.paginas,
                                                                   'editorcompilador': capitulos.editorcompilador,
                                                                   'totalcapitulo': capitulos.totalcapitulo,
                                                                   'fechapublicacion': capitulos.fechapublicacion,
                                                                   'filiacion': capitulos.filiacion,
                                                                   'areaconocimiento': capitulos.areaconocimiento,
                                                                   'subareaconocimiento': capitulos.subareaconocimiento,
                                                                   'subareaespecificaconocimiento': capitulos.subareaespecificaconocimiento,
                                                                   'lineainvestigacion': capitulos.lineainvestigacion,
                                                                   'sublineainvestigacion': capitulos.sublineainvestigacion,
                                                                   'provieneproyecto': capitulos.provieneproyecto,
                                                                   'tipoproyecto': capitulos.tipoproyecto,
                                                                   'proyectointerno': capitulos.proyectointerno,
                                                                   'proyectoexterno': capitulos.proyectoexterno,
                                                                   'pertenecegrupoinv': capitulos.pertenecegrupoinv,
                                                                   'grupoinvestigacion': capitulos.grupoinvestigacion,
                                                                   'revisadopar': capitulos.revisadopar
                                                                   })
                    form.editar(capitulos)
                    data['form'] = form
                    return render(request, "inv_libros/editarcapitulo.html", data)
                except Exception as ex:
                    pass

            elif action == 'deleteparticipantelibro':
                try:
                    data['title'] = u'Eliminar Participante'
                    tipo = request.GET['tipo']
                    data['participante'] = participante = ParticipanteLibros.objects.get(pk=request.GET['id'])
                    if tipo == '1':
                        data['nombres'] = participante.profesor.persona.nombre_completo()
                    elif tipo == '3':
                        data['nombres'] = participante.administrativo.persona.nombre_completo()
                    else:
                        data['nombres'] = participante.inscripcion.persona.nombre_completo()
                    return render(request, "inv_libros/deleteparticipantelibro.html", data)
                except Exception as ex:
                    pass

            elif action == 'deleteparticipantecapitulo':
                try:
                    data['title'] = u'Eliminar Participante'
                    tipo = request.GET['tipo']
                    data['participante'] = participante = ParticipanteCapituloLibros.objects.get(pk=request.GET['id'])
                    if tipo == '1':
                        data['nombres'] = participante.profesor.persona.nombre_completo()
                    if tipo == '3':
                        data['nombres'] = participante.administrativo.persona.nombre_completo()
                    if tipo == '2':
                        data['nombres'] = participante.inscripcion.persona.nombre_completo()
                    return render(request, "inv_libros/deleteparticipantecapitulo.html", data)
                except Exception as ex:
                    pass

            elif action == 'deletelibro':
                try:
                    data['title'] = u'Eliminar Libro'
                    data['libros'] = LibroInvestigacion.objects.get(pk=request.GET['idlibro'])
                    return render(request, "inv_libros/deletelibro.html", data)
                except Exception as ex:
                    pass

            elif action == 'deletecapitulo':
                try:
                    data['title'] = u'Eliminar Capitulo'
                    data['capitulos'] = CapituloLibroInvestigacion.objects.get(pk=request.GET['idcapitulo'])
                    return render(request, "inv_libros/deletecapitulo.html", data)
                except Exception as ex:
                    pass

            elif action == 'participanteslibro':
                try:
                    data['title'] = u'Participantes de Libros'
                    data['libros'] = libros = LibroInvestigacion.objects.get(pk=int(encrypt(request.GET['id'])))
                    data['tipoparticipante'] = TIPO_PARTICIPANTE
                    data['tipoparinstitucion'] = TIPO_PARTICIPANTE_INSTITUCION
                    data['participantes'] = ParticipanteLibros.objects.filter(status=True,libros=libros)
                    return render(request, "inv_libros/participanteslibro.html", data)

                except Exception as ex:
                    pass

            elif action == 'participantescapitulos':
                try:
                    data['title'] = u'Participantes de Capítulos'
                    data['capitulo'] = capitulo = CapituloLibroInvestigacion.objects.get(pk=int(encrypt(request.GET['id'])))
                    data['tipoparinstitucion'] = TIPO_PARTICIPANTE_INSTITUCION
                    data['participantes'] = ParticipanteCapituloLibros.objects.filter(status=True, matrizevidencia_id=7, capitulolibros=capitulo)
                    return render(request, "inv_libros/participantescapitulos.html", data)
                except Exception as ex:
                    pass

            elif action == 'addparticipantesdocentes':
                try:
                    data['title'] = u'Participante Docente'
                    data['form'] = ParticipanteProfesorPonenciaForm()
                    data['id'] = request.GET['idlibro']
                    return render(request, "inv_libros/addparticipantedocente.html", data)
                except Exception as ex:
                    pass

            elif action == 'addparticipantescapitulodocentes':
                try:
                    data['title'] = u'Participante Docente'
                    data['form'] = ParticipanteProfesorCapituloLibroForm
                    data['id'] = request.GET['idcapitulo']
                    return render(request, "inv_libros/addparticipantedocentecapitu.html", data)
                except Exception as ex:
                    pass

            elif action == 'addparticipantesadministrativos':
                try:
                    data['title'] = u'Participante Administrativo'
                    data['form'] = ParticipanteAdministrativoLibroForm()
                    data['id'] = request.GET['idlibro']
                    data['tipoparinstitucion'] = TIPO_PARTICIPANTE_INSTITUCION
                    return render(request, "inv_libros/addparticipanteadministrativo.html", data)
                except Exception as ex:
                    pass

            elif action == 'addparticipantescapituloadministrativos':
                try:
                    data['title'] = u'Participante Administrativo'
                    data['form'] = ParticipanteAdministrativoCapituloLibroForm
                    data['id'] = request.GET['idcapitulo']
                    return render(request, "inv_libros/addparticipantecapituloadministrativo.html", data)
                except Exception as ex:
                    pass

            elif action == 'evidenciaslibro':
                try:
                    data['title'] = u'Evidencia Libro'
                    data['libros'] = LibroInvestigacion.objects.get(pk=int(encrypt(request.GET['id'])))
                    data['evidencias'] = Evidencia.objects.filter(status=True, matrizevidencia_id=6)
                    data['formevidencias'] = EvidenciaForm()
                    return render(request, "inv_libros/evidenciaslibros.html", data)
                except Exception as ex:
                    pass

            elif action == 'evidenciascapitulo':
                try:
                    data['title'] = u'Evidencia Capitulo'
                    data['capitulos'] = CapituloLibroInvestigacion.objects.get(pk=int(encrypt(request.GET['id'])))
                    data['evidencias'] = Evidencia.objects.filter(status=True, matrizevidencia_id=7)
                    data['formevidencias'] = EvidenciaForm()
                    return render(request, "inv_libros/evidenciascapitulos.html", data)
                except Exception as ex:
                    pass

            elif action == 'addevidenciaslibro':
                try:
                    evidencia = Evidencia.objects.get(pk=int(encrypt(request.GET['idevidencia'])))
                    detalleevidencia = DetalleEvidencias.objects.filter(evidencia_id=int(encrypt(request.GET['idevidencia'])), libro_id=int(encrypt(request.GET['id'])))
                    data['title'] = u'Evidencia Libros'
                    data['evidencia'] = evidencia.__str__().upper()

                    if detalleevidencia:
                        detalleevidencia = detalleevidencia[0]
                        form = EvidenciaForm(initial={'descripcion': detalleevidencia.descripcion})
                    else:
                        form = EvidenciaForm(initial={'descripcion': evidencia.nombre.upper()})

                    data['form'] = form
                    data['id'] = request.GET['id']
                    data['idevidencia'] = request.GET['idevidencia']

                    template = get_template("inv_libros/add_evidenciaslibros.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content})
                except Exception as ex:
                    pass

            elif action == 'addevidenciascapitulo':
                try:
                    evidencia = Evidencia.objects.get(pk=int(encrypt(request.GET['idevidencia'])))
                    detalleevidencia = DetalleEvidencias.objects.filter(evidencia_id=int(encrypt(request.GET['idevidencia'])), capitulo_id=int(encrypt(request.GET['id'])))
                    data['title'] = u'Evidencia Capitulos'
                    data['evidencia'] = evidencia.__str__().upper()
                    if detalleevidencia:
                        detalleevidencia = detalleevidencia[0]
                        form = EvidenciaForm(initial={'descripcion': detalleevidencia.descripcion})
                    else:
                        form = EvidenciaForm(initial={'descripcion': evidencia.nombre.upper()})

                    data['form'] = form
                    data['id'] = request.GET['id']
                    data['idevidencia'] = request.GET['idevidencia']
                    template = get_template("inv_libros/add_evidenciascapitulos.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content})
                except Exception as ex:
                    pass

            elif action == 'listacapitulos':
                try:
                    data['title'] = u'Listado de Capítulos'
                    search = None
                    ids = None
                    tipobus = 0 if 'tipobus' not in request.GET else int(request.GET['tipobus'])

                    if 'id' in request.GET:
                        tipobus = 1
                        ids = request.GET['id']
                        capituloslibros = CapituloLibroInvestigacion.objects.select_related().filter(Q(status=True) | Q(status=False, eliminadoxdoc=True), pk=int(encrypt(ids))).order_by('titulocapitulo')
                    elif 's' in request.GET:
                        search = request.GET['s']
                        if search.isdigit():
                            if tipobus != 4:
                                capituloslibros = CapituloLibroInvestigacion.objects.select_related().filter(Q(status=True) | Q(status=False, eliminadoxdoc=True), pk=search)
                            else:
                                anio = int(search)
                                capituloslibros = CapituloLibroInvestigacion.objects.select_related().filter(Q(status=True) | Q(status=False, eliminadoxdoc=True), fechapublicacion__year=anio)#.order_by('titulocapitulo')
                        else:
                            if tipobus == 1:
                                capituloslibros = CapituloLibroInvestigacion.objects.select_related().filter(Q(status=True) | Q(status=False, eliminadoxdoc=True), titulocapitulo__icontains=search)
                            elif tipobus == 2:
                                capituloslibros = CapituloLibroInvestigacion.objects.select_related().filter(Q(status=True) | Q(status=False, eliminadoxdoc=True), titulolibro__icontains=search)
                            else:
                                ss = search.split(' ')
                                if len(ss) == 1:
                                    capituloslibros = CapituloLibroInvestigacion.objects.select_related().filter(Q(participantecapitulolibros__profesor__persona__nombres__icontains=search) |
                                                                                                                 Q(participantecapitulolibros__profesor__persona__apellido1__icontains=search) |
                                                                                                                 Q(participantecapitulolibros__profesor__persona__apellido2__icontains=search) |
                                                                                                                 Q(participantecapitulolibros__administrativo__persona__nombres__icontains=search) |
                                                                                                                 Q(participantecapitulolibros__administrativo__persona__apellido1__icontains=search) |
                                                                                                                 Q(participantecapitulolibros__administrativo__persona__apellido2__icontains=search)
                                                                                                                 ,Q(status=True) | Q(status=False, eliminadoxdoc=True))
                                else:
                                    capituloslibros = CapituloLibroInvestigacion.objects.select_related().filter(

                                        (Q(participantecapitulolibros__profesor__persona__apellido1__icontains=ss[0]) &
                                        Q(participantecapitulolibros__profesor__persona__apellido2__icontains=ss[1])) |

                                        (Q(participantecapitulolibros__administrativo__persona__apellido1__icontains=ss[0]) &
                                        Q(participantecapitulolibros__administrativo__persona__apellido2__icontains=ss[1]))
                                        , Q(status=True) | Q(status=False, eliminadoxdoc=True))
                    elif 'idsp' in request.GET:
                        tipobus = 3
                        lista_ids_personas = request.GET['idsp'].split(",")

                        capituloslibros = CapituloLibroInvestigacion.objects.select_related().filter(Q(participantecapitulolibros__profesor__persona__id__in=lista_ids_personas) |
                                                                                                     Q(participantecapitulolibros__administrativo__persona__id__in=lista_ids_personas),
                                                                                                     Q(status=True) | Q(status=False, eliminadoxdoc=True), participantecapitulolibros__status=True)


                    else:
                        tipobus = 1
                        capituloslibros = CapituloLibroInvestigacion.objects.select_related().filter(Q(status=True) | Q(status=False, eliminadoxdoc=True))#.order_by('titulocapitulo')

                    capituloslibros = capituloslibros.order_by('-fechapublicacion')

                    estadocapitulo = 0
                    if 'estadocapitulo' in request.GET:
                        estadocapitulo = int(request.GET['estadocapitulo'])
                        if estadocapitulo > 0:
                            if estadocapitulo == 1:
                                capituloslibros = capituloslibros.filter(aprobado=True)
                            else:
                                capituloslibros = capituloslibros.filter(aprobado=False)

                    paging = MiPaginador(capituloslibros, 25)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['page'] = page
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['listacapitulos'] = page.object_list
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['tipobus'] = tipobus
                    data['totalcapitulos'] = total = capituloslibros.count()
                    data['totalaprobados'] = totalaprobados = capituloslibros.filter(aprobado=True).count()
                    data['totalporaprobar'] = total - totalaprobados
                    data['estadocapitulo'] = estadocapitulo
                    data['aniosevidencias'] = CapituloLibroInvestigacion.objects.filter(Q(status=True) | Q(status=False, eliminadoxdoc=True), fechapublicacion__isnull=False).annotate(anio=ExtractYear('fechapublicacion')).values_list('anio', flat=True).order_by('-anio').distinct()

                    data['participantes'] = Persona.objects.values('id', 'cedula', 'nombres', 'apellido1', 'apellido2').filter(Q(profesor__participantecapitulolibros__isnull=False, profesor__participantecapitulolibros__status=True))

                    return render(request, "inv_libros/viewcapitulos.html", data)
                except Exception as ex:
                    pass

            elif action == 'reporte_libros_excel':
                try:
                    __author__ = 'Unemi'
                    fuentenormal = easyxtitulo = easyxf(
                        'font: name Times New Roman, color-index black, bold on , height 350; alignment: horiz centre')
                    titulo = easyxf(
                        'font: name Times New Roman, color-index black, bold on , height 350; alignment: horiz centre')
                    titulo2 = easyxf(
                        'font: name Times New Roman, color-index black, bold on , height 250; alignment: horiz centre')
                    fuentecabecera = easyxf(
                        'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
                    fuentenormal = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
                    fuentenormalwrap = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
                    fuentenormalwrap.alignment.wrap = True
                    fuentenormalcent = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz centre')
                    fuentemoneda = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right',
                        num_format_str=' "$" #,##0.00')
                    fuentefecha = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz center',
                        num_format_str='yyyy-mm-dd')
                    fuentenumerodecimal = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right',
                        num_format_str='#,##0.00')
                    fuentenumeroentero = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right')

                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('Listado')
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=lista_libros_' + random.randint(1, 10000).__str__() + '.xls'

                    ws.write_merge(0, 0, 0, 21, 'UNIVERSIDAD ESTATAL DE MILAGRO', titulo)
                    ws.write_merge(1, 1, 0, 21, 'VICERRECTORADO DE INVESTIGACIÓN Y POSGRADO', titulo2)
                    ws.write_merge(2, 2, 0, 21, 'COORDINACIÓN DE INVESTIGACIÓN', titulo2)
                    ws.write_merge(3, 3, 0, 21, 'LISTADO DE LIBROS', titulo2)

                    row_num = 5
                    columns = [
                        (u"TIPO DE ARTÍCULOS", 2500),
                        (u"CÓDIGO", 4500),
                        (u"NOMBRE DEL LIBRO", 10000),
                        (u"CÓDIGO ISBN", 5000),
                        (u"REVISADA POR PARES", 3000),
                        (u"FILIACIÓN", 3000),
                        (u"FECHA PUBLICACIÓN", 3000),
                        (u"ÁREA DE CONOCIMIENTO", 10000),
                        (u"SUB-ÁREA DE CONOCIMIENTO", 10000),
                        (u"SUB-ÁREA ESPECÍFICA DE CONOCIMIENTO", 10000),
                        (u"LÍNEA INVESTIGACIÓN", 10000),
                        (u"SUB-LÍNEA INVESTIGACIÓN", 10000),
                        (u"PROVIENE DE PROYECTO", 3000),
                        (u"TIPO PROYECTO", 3000),
                        (u"TÍTULO DEL PROYECTO", 10000),
                        (u"PERTENECE GRUPO INVESTIGACIÓN", 3000),
                        (u"GRUPO DE INVESTIGACIÓN", 10000),
                        (u"CÉDULA", 3000),
                        (u"TIPO PARTICIPANTE", 3000),
                        (u"PARTICIPANTE", 15000),
                        (u"TIPO PARTICIPACIÓN", 3000),
                        (u"TIPO UNEMI", 3000)
                    ]

                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], fuentecabecera)
                        ws.col(col_num).width = columns[col_num][1]

                    listalibros = ParticipanteLibros.objects.filter(libros__status=True, status=True).order_by('libros__nombrelibro')

                    for libros in listalibros:
                        row_num += 1

                        if libros.profesor:
                            nombres = libros.profesor.persona.nombre_completo_inverso()
                            cedula = libros.profesor.persona.cedula
                        elif libros.administrativo:
                            nombres = libros.administrativo.persona.nombre_completo_inverso()
                            cedula = libros.administrativo.persona.cedula

                        ws.write(row_num, 0, "LIBRO", fuentenormal)
                        ws.write(row_num, 1, libros.libros.codisbn + ' ' + str(libros.libros.id) + '-LIB', fuentenormal)
                        ws.write(row_num, 2, libros.libros.nombrelibro.strip(), fuentenormal)
                        ws.write(row_num, 3, libros.libros.codisbn, fuentenormal)
                        ws.write(row_num, 4, libros.libros.get_revi_pare_display(), fuentenormalcent)
                        ws.write(row_num, 5, libros.libros.get_revi_fili_display(), fuentenormalcent)
                        ws.write(row_num, 6, libros.libros.fechapublicacion, fuentefecha)
                        ws.write(row_num, 7, libros.libros.areaconocimiento.nombre if libros.libros.areaconocimiento else "", fuentenormal)
                        ws.write(row_num, 8, libros.libros.subareaconocimiento.nombre if libros.libros.subareaconocimiento else "", fuentenormal)
                        ws.write(row_num, 9, libros.libros.subareaespecificaconocimiento.nombre if libros.libros.subareaespecificaconocimiento else "", fuentenormal)
                        ws.write(row_num, 10, libros.libros.lineainvestigacion.nombre if libros.libros.lineainvestigacion else '', fuentenormal)
                        ws.write(row_num, 11, libros.libros.sublineainvestigacion.nombre if libros.libros.sublineainvestigacion else '', fuentenormal)
                        ws.write(row_num, 12, 'SI' if libros.libros.tipoproyecto else 'NO', fuentenormal)
                        ws.write(row_num, 13, libros.libros.get_tipoproyecto_display() if libros.libros.tipoproyecto else '', fuentenormal)
                        if libros.libros.tipoproyecto:
                            ws.write(row_num, 14, libros.libros.proyectointerno.nombre if libros.libros.proyectointerno else libros.libros.proyectoexterno.nombre, fuentenormal)
                        else:
                            ws.write(row_num, 14, '', fuentenormal)

                        ws.write(row_num, 15, 'SI' if libros.libros.pertenecegrupoinv else 'NO', fuentenormal)
                        ws.write(row_num, 16, libros.libros.grupoinvestigacion.nombre if libros.libros.pertenecegrupoinv else '', fuentenormal)
                        ws.write(row_num, 17, cedula, fuentenormal)
                        ws.write(row_num, 18, libros.get_tipoparticipante_display(), fuentenormal)
                        ws.write(row_num, 19, nombres, fuentenormal)
                        ws.write(row_num, 20, libros.get_tipoparticipanteins_display(), fuentenormal)
                        ws.write(row_num, 21, libros.tipo_unemi(), fuentenormal)


                    wb.save(response)
                    return response
                except Exception as ex:
                    pass

            elif action == 'reporte_libros_excel_participante':
                try:
                    __author__ = 'Unemi'
                    fuentenormal = easyxtitulo = easyxf(
                        'font: name Times New Roman, color-index black, bold on , height 350; alignment: horiz centre')
                    titulo = easyxf(
                        'font: name Times New Roman, color-index black, bold on , height 350; alignment: horiz centre')
                    titulo2 = easyxf(
                        'font: name Times New Roman, color-index black, bold on , height 250; alignment: horiz centre')
                    fuentecabecera = easyxf(
                        'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
                    fuentenormal = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
                    fuentenormalwrap = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
                    fuentenormalwrap.alignment.wrap = True
                    fuentenormalcent = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz centre')
                    fuentemoneda = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right',
                        num_format_str=' "$" #,##0.00')
                    fuentefecha = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz center',
                        num_format_str='yyyy-mm-dd')
                    fuentenumerodecimal = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right',
                        num_format_str='#,##0.00')
                    fuentenumeroentero = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right')

                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('Listado')
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=lista_libros_participante_' + random.randint(1, 10000).__str__() + '.xls'

                    ws.write_merge(0, 0, 0, 21, 'UNIVERSIDAD ESTATAL DE MILAGRO', titulo)
                    ws.write_merge(1, 1, 0, 21, 'VICERRECTORADO DE INVESTIGACIÓN Y POSGRADO', titulo2)
                    ws.write_merge(2, 2, 0, 21, 'COORDINACIÓN DE INVESTIGACIÓN', titulo2)
                    ws.write_merge(3, 3, 0, 21, 'LISTADO DE LIBROS POR PARTICIPANTE', titulo2)

                    row_num = 5
                    columns = [
                        (u"TIPO DE ARTÍCULOS", 2500),
                        (u"CÓDIGO", 4500),
                        (u"NOMBRE DEL LIBRO", 10000),
                        (u"CÓDIGO ISBN", 5000),
                        (u"REVISADA POR PARES", 3000),
                        (u"FILIACIÓN", 3000),
                        (u"FECHA PUBLICACIÓN", 3000),
                        (u"ÁREA DE CONOCIMIENTO", 10000),
                        (u"SUB-ÁREA DE CONOCIMIENTO", 10000),
                        (u"SUB-ÁREA ESPECÍFICA DE CONOCIMIENTO", 10000),
                        (u"LÍNEA INVESTIGACIÓN", 10000),
                        (u"SUB-LÍNEA INVESTIGACIÓN", 10000),
                        (u"PROVIENE DE PROYECTO", 3000),
                        (u"TIPO PROYECTO", 3000),
                        (u"TÍTULO DEL PROYECTO", 10000),
                        (u"PERTENECE GRUPO INVESTIGACIÓN", 3000),
                        (u"GRUPO DE INVESTIGACIÓN", 10000),
                        (u"CÉDULA", 3000),
                        (u"TIPO PARTICIPANTE", 3000),
                        (u"PARTICIPANTE", 15000),
                        (u"TIPO PARTICIPACIÓN", 3000),
                        (u"TIPO UNEMI", 3000)
                    ]

                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], fuentecabecera)
                        ws.col(col_num).width = columns[col_num][1]

                    lista_ids_personas = request.GET['idsp'].split(",")

                    listalibros = ParticipanteLibros.objects.filter(Q(profesor__persona__id__in=lista_ids_personas)|
                                                                    Q(administrativo__persona__id__in=lista_ids_personas),
                                                                    libros__status=True, status=True).order_by('libros__nombrelibro')

                    for libros in listalibros:
                        row_num += 1

                        if libros.profesor:
                            nombres = libros.profesor.persona.nombre_completo_inverso()
                            cedula = libros.profesor.persona.cedula
                        elif libros.administrativo:
                            nombres = libros.administrativo.persona.nombre_completo_inverso()
                            cedula = libros.administrativo.persona.cedula

                        ws.write(row_num, 0, "LIBRO", fuentenormal)
                        ws.write(row_num, 1, libros.libros.codisbn + ' ' + str(libros.libros.id) + '-LIB', fuentenormal)
                        ws.write(row_num, 2, libros.libros.nombrelibro.strip(), fuentenormal)
                        ws.write(row_num, 3, libros.libros.codisbn, fuentenormal)
                        ws.write(row_num, 4, libros.libros.get_revi_pare_display(), fuentenormalcent)
                        ws.write(row_num, 5, libros.libros.get_revi_fili_display(), fuentenormalcent)
                        ws.write(row_num, 6, libros.libros.fechapublicacion, fuentefecha)
                        ws.write(row_num, 7, libros.libros.areaconocimiento.nombre if libros.libros.areaconocimiento else "", fuentenormal)
                        ws.write(row_num, 8, libros.libros.subareaconocimiento.nombre if libros.libros.subareaconocimiento else "", fuentenormal)
                        ws.write(row_num, 9, libros.libros.subareaespecificaconocimiento.nombre if libros.libros.subareaespecificaconocimiento else "", fuentenormal)
                        ws.write(row_num, 10, libros.libros.lineainvestigacion.nombre if libros.libros.lineainvestigacion else '', fuentenormal)
                        ws.write(row_num, 11, libros.libros.sublineainvestigacion.nombre if libros.libros.sublineainvestigacion else '', fuentenormal)
                        ws.write(row_num, 12, 'SI' if libros.libros.tipoproyecto else 'NO', fuentenormal)
                        ws.write(row_num, 13, libros.libros.get_tipoproyecto_display() if libros.libros.tipoproyecto else '', fuentenormal)
                        if libros.libros.tipoproyecto:
                            ws.write(row_num, 14, libros.libros.proyectointerno.nombre if libros.libros.proyectointerno else libros.libros.proyectoexterno.nombre, fuentenormal)
                        else:
                            ws.write(row_num, 14, '', fuentenormal)

                        ws.write(row_num, 15, 'SI' if libros.libros.pertenecegrupoinv else 'NO', fuentenormal)
                        ws.write(row_num, 16, libros.libros.grupoinvestigacion.nombre if libros.libros.pertenecegrupoinv else '', fuentenormal)
                        ws.write(row_num, 17, cedula, fuentenormal)
                        ws.write(row_num, 18, libros.get_tipoparticipante_display(), fuentenormal)
                        ws.write(row_num, 19, nombres, fuentenormal)
                        ws.write(row_num, 20, libros.get_tipoparticipanteins_display(), fuentenormal)
                        ws.write(row_num, 21, libros.tipo_unemi(), fuentenormal)

                    wb.save(response)
                    return response
                except Exception as ex:
                    pass

            elif action == 'reporte_capitulos_libro_excel':
                try:
                    __author__ = 'Unemi'
                    fuentenormal = easyxtitulo = easyxf(
                        'font: name Times New Roman, color-index black, bold on , height 350; alignment: horiz centre')
                    titulo = easyxf(
                        'font: name Times New Roman, color-index black, bold on , height 350; alignment: horiz centre')
                    titulo2 = easyxf(
                        'font: name Times New Roman, color-index black, bold on , height 250; alignment: horiz centre')
                    fuentecabecera = easyxf(
                        'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
                    fuentenormal = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
                    fuentenormalwrap = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
                    fuentenormalwrap.alignment.wrap = True
                    fuentenormalcent = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz centre')
                    fuentemoneda = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right',
                        num_format_str=' "$" #,##0.00')
                    fuentefecha = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz center',
                        num_format_str='yyyy-mm-dd')
                    fuentenumerodecimal = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right',
                        num_format_str='#,##0.00')
                    fuentenumeroentero = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right')

                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('Listado')
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=lista_capitulos_libro_' + random.randint(1, 10000).__str__() + '.xls'

                    ws.write_merge(0, 0, 0, 22, 'UNIVERSIDAD ESTATAL DE MILAGRO', titulo)
                    ws.write_merge(1, 1, 0, 22, 'VICERRECTORADO DE INVESTIGACIÓN Y POSGRADO', titulo2)
                    ws.write_merge(2, 2, 0, 22, 'COORDINACIÓN DE INVESTIGACIÓN', titulo2)
                    ws.write_merge(3, 3, 0, 22, 'LISTADO DE CAPÍTULOS DE LIBROS', titulo2)

                    row_num = 5

                    columns = [
                        (u"TIPO DE ARTÍCULOS", 2500),
                        (u"CÓDIGO", 4500),
                        (u"NOMBRE DEL LIBRO", 10000),
                        (u"NOMBRE DEL CAPÍTULO", 10000),
                        (u"CODIGO ISBN", 5000),
                        (u"EDITOR O COMPILADOR", 5000),
                        (u"PÁGINAS", 5000),
                        (u"FECHA PUBLICACIÓN", 3000),
                        (u"ÁREA DE CONOCIMIENTO", 10000),
                        (u"SUB-ÁREA DE CONOCIMIENTO", 10000),
                        (u"SUB-ÁREA ESPECÍFICA DE CONOCIMIENTO", 10000),
                        (u"LÍNEA INVESTIGACIÓN", 10000),
                        (u"SUB-LÍNEA INVESTIGACIÓN", 10000),
                        (u"PROVIENE DE PROYECTO", 3000),
                        (u"TIPO PROYECTO", 3000),
                        (u"TÍTULO DEL PROYECTO", 10000),
                        (u"PERTENECE GRUPO INVESTIGACIÓN", 3000),
                        (u"GRUPO DE INVESTIGACIÓN", 10000),
                        (u"CÉDULA", 3000),
                        (u"PARTICIPANTE", 15000),
                        (u"TIPO PARTICIPANTE", 3000),
                        (u"TIPO UNEMI", 3000),
                        (u"TIPO FILIACION", 3000)
                    ]

                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], fuentecabecera)
                        ws.col(col_num).width = columns[col_num][1]

                    listalibros = ParticipanteCapituloLibros.objects.filter(capitulolibros__status=True, status=True).order_by('capitulolibros__titulocapitulo')

                    for libros in listalibros:
                        row_num += 1

                        if libros.profesor:
                            nombres = libros.profesor.persona.nombre_completo_inverso()
                            cedula = libros.profesor.persona.cedula
                        elif libros.administrativo:
                            nombres = libros.administrativo.persona.nombre_completo_inverso()
                            cedula = libros.administrativo.persona.cedula

                        ws.write(row_num, 0, "CAPITULO LIBRO", fuentenormal)
                        ws.write(row_num, 1, libros.capitulolibros.codisbn + ' ' + str(libros.capitulolibros.id) + '-CAPLIB', fuentenormal)
                        ws.write(row_num, 2, libros.capitulolibros.titulolibro.strip(), fuentenormal)
                        ws.write(row_num, 3, libros.capitulolibros.titulocapitulo.strip(), fuentenormal)
                        ws.write(row_num, 4, libros.capitulolibros.codisbn, fuentenormal)
                        ws.write(row_num, 5, libros.capitulolibros.editorcompilador.strip(), fuentenormal)
                        ws.write(row_num, 6, libros.capitulolibros.paginas, fuentenormal)
                        ws.write(row_num, 7, libros.capitulolibros.fechapublicacion, fuentefecha)
                        ws.write(row_num, 8, libros.capitulolibros.areaconocimiento.nombre if libros.capitulolibros.areaconocimiento else "", fuentenormal)
                        ws.write(row_num, 9, libros.capitulolibros.subareaconocimiento.nombre if libros.capitulolibros.subareaconocimiento.nombre else "", fuentenormal)
                        ws.write(row_num, 10, libros.capitulolibros.subareaespecificaconocimiento.nombre if libros.capitulolibros.subareaespecificaconocimiento else "", fuentenormal)
                        ws.write(row_num, 11, libros.capitulolibros.lineainvestigacion.nombre if libros.capitulolibros.lineainvestigacion else '', fuentenormal)
                        ws.write(row_num, 12, libros.capitulolibros.sublineainvestigacion.nombre if libros.capitulolibros.sublineainvestigacion else '', fuentenormal)
                        ws.write(row_num, 13, 'SI' if libros.capitulolibros.tipoproyecto else 'NO', fuentenormal)
                        ws.write(row_num, 14, libros.capitulolibros.get_tipoproyecto_display() if libros.capitulolibros.tipoproyecto else '', fuentenormal)
                        if libros.capitulolibros.tipoproyecto:
                            ws.write(row_num, 15, libros.capitulolibros.proyectointerno.nombre if libros.capitulolibros.proyectointerno else libros.capitulolibros.proyectoexterno.nombre, fuentenormal)
                        else:
                            ws.write(row_num, 15, '', fuentenormal)

                        ws.write(row_num, 16, 'SI' if libros.capitulolibros.pertenecegrupoinv else 'NO', fuentenormal)
                        ws.write(row_num, 17, libros.capitulolibros.grupoinvestigacion.nombre if libros.capitulolibros.pertenecegrupoinv else '', fuentenormal)
                        ws.write(row_num, 18, cedula, fuentenormal)
                        ws.write(row_num, 19, nombres, fuentenormal)
                        ws.write(row_num, 20, libros.get_tipoparticipante_display(), fuentenormal)
                        ws.write(row_num, 21, libros.tipo_unemi(), fuentenormal)
                        ws.write(row_num, 22, libros.get_tipoparticipanteins_display(), fuentenormal)

                    wb.save(response)
                    return response
                except Exception as ex:
                    pass

            elif action == 'reporte_capitulos_libro_excel_participante':
                try:
                    __author__ = 'Unemi'
                    fuentenormal = easyxtitulo = easyxf(
                        'font: name Times New Roman, color-index black, bold on , height 350; alignment: horiz centre')
                    titulo = easyxf(
                        'font: name Times New Roman, color-index black, bold on , height 350; alignment: horiz centre')
                    titulo2 = easyxf(
                        'font: name Times New Roman, color-index black, bold on , height 250; alignment: horiz centre')
                    fuentecabecera = easyxf(
                        'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
                    fuentenormal = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
                    fuentenormalwrap = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
                    fuentenormalwrap.alignment.wrap = True
                    fuentenormalcent = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz centre')
                    fuentemoneda = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right',
                        num_format_str=' "$" #,##0.00')
                    fuentefecha = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz center',
                        num_format_str='yyyy-mm-dd')
                    fuentenumerodecimal = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right',
                        num_format_str='#,##0.00')
                    fuentenumeroentero = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right')

                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('Listado')
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=lista_capitulos_libro_participante_' + random.randint(1, 10000).__str__() + '.xls'

                    ws.write_merge(0, 0, 0, 22, 'UNIVERSIDAD ESTATAL DE MILAGRO', titulo)
                    ws.write_merge(1, 1, 0, 22, 'VICERRECTORADO DE INVESTIGACIÓN Y POSGRADO', titulo2)
                    ws.write_merge(2, 2, 0, 22, 'COORDINACIÓN DE INVESTIGACIÓN', titulo2)
                    ws.write_merge(3, 3, 0, 22, 'LISTADO DE CAPÍTULOS DE LIBROS POR PARTICIPANTE', titulo2)

                    row_num = 5

                    columns = [
                        (u"TIPO DE ARTÍCULOS", 2500),
                        (u"CÓDIGO", 4500),
                        (u"NOMBRE DEL LIBRO", 10000),
                        (u"NOMBRE DEL CAPÍTULO", 10000),
                        (u"CODIGO ISBN", 5000),
                        (u"EDITOR O COMPILADOR", 5000),
                        (u"PÁGINAS", 5000),
                        (u"FECHA PUBLICACIÓN", 3000),
                        (u"ÁREA DE CONOCIMIENTO", 10000),
                        (u"SUB-ÁREA DE CONOCIMIENTO", 10000),
                        (u"SUB-ÁREA ESPECÍFICA DE CONOCIMIENTO", 10000),
                        (u"LÍNEA INVESTIGACIÓN", 10000),
                        (u"SUB-LÍNEA INVESTIGACIÓN", 10000),
                        (u"PROVIENE DE PROYECTO", 3000),
                        (u"TIPO PROYECTO", 3000),
                        (u"TÍTULO DEL PROYECTO", 10000),
                        (u"PERTENECE GRUPO INVESTIGACIÓN", 3000),
                        (u"GRUPO DE INVESTIGACIÓN", 10000),
                        (u"CÉDULA", 3000),
                        (u"PARTICIPANTE", 15000),
                        (u"TIPO PARTICIPANTE", 3000),
                        (u"TIPO UNEMI", 3000),
                        (u"TIPO FILIACION", 3000)
                    ]

                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], fuentecabecera)
                        ws.col(col_num).width = columns[col_num][1]

                    lista_ids_personas = request.GET['idsp'].split(",")

                    listalibros = ParticipanteCapituloLibros.objects.filter(Q(profesor__persona__id__in=lista_ids_personas) |
                                                                            Q(administrativo__persona__id__in=lista_ids_personas),
                                                                            capitulolibros__status=True, status=True).order_by('capitulolibros__titulocapitulo')

                    for libros in listalibros:
                        row_num += 1

                        if libros.profesor:
                            nombres = libros.profesor.persona.nombre_completo_inverso()
                            cedula = libros.profesor.persona.cedula
                        elif libros.administrativo:
                            nombres = libros.administrativo.persona.nombre_completo_inverso()
                            cedula = libros.administrativo.persona.cedula

                        ws.write(row_num, 0, "CAPITULO LIBRO", fuentenormal)
                        ws.write(row_num, 1, libros.capitulolibros.codisbn + ' ' + str(libros.capitulolibros.id) + '-CAPLIB', fuentenormal)
                        ws.write(row_num, 2, libros.capitulolibros.titulolibro.strip(), fuentenormal)
                        ws.write(row_num, 3, libros.capitulolibros.titulocapitulo.strip(), fuentenormal)
                        ws.write(row_num, 4, libros.capitulolibros.codisbn, fuentenormal)
                        ws.write(row_num, 5, libros.capitulolibros.editorcompilador.strip(), fuentenormal)
                        ws.write(row_num, 6, libros.capitulolibros.paginas, fuentenormal)
                        ws.write(row_num, 7, libros.capitulolibros.fechapublicacion, fuentefecha)
                        ws.write(row_num, 8, libros.capitulolibros.areaconocimiento.nombre if libros.capitulolibros.areaconocimiento else "", fuentenormal)
                        ws.write(row_num, 9, libros.capitulolibros.subareaconocimiento.nombre if libros.capitulolibros.subareaconocimiento.nombre else "", fuentenormal)
                        ws.write(row_num, 10, libros.capitulolibros.subareaespecificaconocimiento.nombre if libros.capitulolibros.subareaespecificaconocimiento else "", fuentenormal)
                        ws.write(row_num, 11, libros.capitulolibros.lineainvestigacion.nombre if libros.capitulolibros.lineainvestigacion else '', fuentenormal)
                        ws.write(row_num, 12, libros.capitulolibros.sublineainvestigacion.nombre if libros.capitulolibros.sublineainvestigacion else '', fuentenormal)
                        ws.write(row_num, 13, 'SI' if libros.capitulolibros.tipoproyecto else 'NO', fuentenormal)
                        ws.write(row_num, 14, libros.capitulolibros.get_tipoproyecto_display() if libros.capitulolibros.tipoproyecto else '', fuentenormal)
                        if libros.capitulolibros.tipoproyecto:
                            ws.write(row_num, 15, libros.capitulolibros.proyectointerno.nombre if libros.capitulolibros.proyectointerno else libros.capitulolibros.proyectoexterno.nombre, fuentenormal)
                        else:
                            ws.write(row_num, 15, '', fuentenormal)

                        ws.write(row_num, 16, 'SI' if libros.capitulolibros.pertenecegrupoinv else 'NO', fuentenormal)
                        ws.write(row_num, 17, libros.capitulolibros.grupoinvestigacion.nombre if libros.capitulolibros.pertenecegrupoinv else '', fuentenormal)
                        ws.write(row_num, 18, cedula, fuentenormal)
                        ws.write(row_num, 19, nombres, fuentenormal)
                        ws.write(row_num, 20, libros.get_tipoparticipante_display(), fuentenormal)
                        ws.write(row_num, 21, libros.tipo_unemi(), fuentenormal)
                        ws.write(row_num, 22, libros.get_tipoparticipanteins_display(), fuentenormal)

                    wb.save(response)
                    return response
                except Exception as ex:
                    pass

            # elif action == 'excelibros':
            #     try:
            #         __author__ = 'Unemi'
            #         style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
            #         style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
            #         style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
            #         title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
            #         style1 = easyxf(num_format_str='D-MMM-YY')
            #         font_style = XFStyle()
            #         font_style.font.bold = True
            #         font_style2 = XFStyle()
            #         font_style2.font.bold = False
            #         wb = Workbook(encoding='utf-8')
            #         ws = wb.add_sheet('exp_xls_post_part')
            #         ws.write_merge(0, 0, 0, 13, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
            #         response = HttpResponse(content_type="application/ms-excel")
            #         response[
            #             'Content-Disposition'] = 'attachment; filename=Listas_Libros' + random.randint(
            #             1, 10000).__str__() + '.xls'
            #
            #         columns = [
            #             (u"TIPO DE ARTICULOS", 2500),
            #             (u"CODIGO", 4500),
            #             (u"NOMBRE DEL LIBRO", 10000),
            #             (u"CODIGO ISBN", 5000),
            #             (u"REVISADA POR PARES", 3000),
            #             (u"FILIACION", 3000),
            #             (u"FECHA PUBLICACION", 3000),
            #             (u"AREA DE CONOCIMIENTO", 10000),
            #             (u"SUBAREA DE CONOCIMIENTO", 10000),
            #             (u"CEDULA", 3000),
            #             (u"TIPO PARTICIPANTE", 3000),
            #             (u"PARTICIPANTE", 15000),
            #             (u"TIPO PARTICIPACIÓN", 3000),
            #             (u"TIPO UNEMI", 3000)
            #         ]
            #         row_num = 3
            #         for col_num in range(len(columns)):
            #             ws.write(row_num, col_num, columns[col_num][0], font_style)
            #             ws.col(col_num).width = columns[col_num][1]
            #         date_format = xlwt.XFStyle()
            #         date_format.num_format_str = 'yyyy/mm/dd'
            #         listalibros = ParticipanteLibros.objects.filter(libros__status=True, status=True).order_by('libros__nombrelibro')
            #         row_num = 4
            #         for libros in listalibros:
            #             i = 0
            #             campo8 = None
            #             campo1 = libros.libros.id
            #             titulolibro = libros.libros.nombrelibro
            #             areaconocimiento = libros.libros.areaconocimiento.nombre
            #             subarea = libros.libros.subareaconocimiento.nombre
            #             tipounemi = libros.tipo_unemi()
            #             tipoparticipante = libros.get_tipoparticipanteins_display()
            #
            #             if libros.profesor:
            #                 nombres = libros.profesor
            #                 cedula = libros.profesor.persona.cedula
            #
            #             if libros.administrativo:
            #                 nombres = libros.administrativo
            #                 cedula = libros.administrativo.persona.cedula
            #
            #             if libros.libros.revi_pare:
            #                 revi_pare = 'SI'
            #             else:
            #                 revi_pare = 'NO'
            #             if libros.libros.revi_fili:
            #                 revi_fili = 'SI'
            #             else:
            #                 revi_fili = 'NO'
            #
            #             ws.write(row_num, 0, 'LIBRO', font_style2)
            #             ws.write(row_num, 1, libros.libros.codisbn + ' ' + str(libros.libros.id) + '-LIB', font_style2)
            #             ws.write(row_num, 2, titulolibro, font_style2)
            #             ws.write(row_num, 3, libros.libros.codisbn, font_style2)
            #             ws.write(row_num, 4, revi_pare, font_style2)
            #             ws.write(row_num, 5, revi_fili, font_style2)
            #             ws.write(row_num, 6, libros.libros.fechapublicacion, date_format)
            #             ws.write(row_num, 7, areaconocimiento, font_style2)
            #             ws.write(row_num, 8, subarea, font_style2)
            #             ws.write(row_num, 9, cedula, font_style2)
            #             ws.write(row_num, 10, libros.get_tipoparticipante_display(), font_style2)
            #             ws.write(row_num, 11, str(nombres), font_style2)
            #             ws.write(row_num, 12, tipoparticipante, font_style2)
            #             ws.write(row_num, 13, tipounemi, font_style2)
            #
            #
            #             row_num += 1
            #         wb.save(response)
            #         return response
            #     except Exception as ex:
            #         pass
            #
            # elif action == 'excelcapitulolibros':
            #     try:
            #         __author__ = 'Unemi'
            #         style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
            #         style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
            #         style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
            #         title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
            #         style1 = easyxf(num_format_str='D-MMM-YY')
            #         font_style = XFStyle()
            #         font_style.font.bold = True
            #         font_style2 = XFStyle()
            #         font_style2.font.bold = False
            #         wb = Workbook(encoding='utf-8')
            #         ws = wb.add_sheet('exp_xls_post_part')
            #         ws.write_merge(0, 0, 0, 11, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
            #         response = HttpResponse(content_type="application/ms-excel")
            #         response[
            #             'Content-Disposition'] = 'attachment; filename=Listas_Capitulo_Libros' + random.randint(
            #             1, 10000).__str__() + '.xls'
            #
            #         columns = [
            #             (u"TIPO DE ARTICULOS", 2500),
            #             (u"CODIGO", 4500),
            #             (u"NOMBRE DEL LIBRO", 10000),
            #             (u"NOMBRE DEL CAPITULO", 10000),
            #             (u"CODIGO ISBN", 5000),
            #             (u"FECHA PUBLICACION", 3000),
            #             (u"AREA DE CONOCIMIENTO", 10000),
            #             (u"SUBAREA DE CONOCIMIENTO", 10000),
            #             (u"CEDULA", 3000),
            #             (u"PARTICIPANTE", 15000),
            #             (u"TIPO PARTICIPANTE", 3000),
            #             (u"TIPO UNEMI", 3000)
            #         ]
            #         row_num = 3
            #         for col_num in range(len(columns)):
            #             ws.write(row_num, col_num, columns[col_num][0], font_style)
            #             ws.col(col_num).width = columns[col_num][1]
            #         date_format = xlwt.XFStyle()
            #         date_format.num_format_str = 'yyyy/mm/dd'
            #         listalibros = ParticipanteCapituloLibros.objects.filter(capitulolibros__status=True, status=True).order_by('capitulolibros__titulocapitulo')
            #         row_num = 4
            #         for libros in listalibros:
            #             i = 0
            #             campo8 = None
            #             campo1 = libros.capitulolibros.id
            #             titulocapitulo = libros.capitulolibros.titulocapitulo
            #
            #             area = libros.capitulolibros.areaconocimiento.nombre
            #             subarea = libros.capitulolibros.subareaconocimiento.nombre
            #             tipounemi = libros.tipo_unemi()
            #             tipoparticipante = libros.get_tipoparticipante_display()
            #
            #             if libros.profesor:
            #                 nombres = libros.profesor
            #                 cedula = libros.profesor.persona.cedula
            #
            #             if libros.administrativo:
            #                 nombres = libros.administrativo
            #                 cedula = libros.administrativo.persona.cedula
            #
            #             ws.write(row_num, 0, 'CAPITULO LIBRO', font_style2)
            #             ws.write(row_num, 1, libros.capitulolibros.codisbn + ' ' + str(campo1) + '-CAPLIB', font_style2)
            #             ws.write(row_num, 2, libros.capitulolibros.titulolibro, font_style2)
            #             ws.write(row_num, 3, titulocapitulo, font_style2)
            #             ws.write(row_num, 4, libros.capitulolibros.codisbn, font_style2)
            #             ws.write(row_num, 5, libros.capitulolibros.fechapublicacion, date_format)
            #             ws.write(row_num, 6, area, font_style2)
            #             ws.write(row_num, 7, subarea, font_style2)
            #             ws.write(row_num, 8, cedula, font_style2)
            #             ws.write(row_num, 9, str(nombres), font_style2)
            #             ws.write(row_num, 10, tipoparticipante, font_style2)
            #             ws.write(row_num, 11, tipounemi, font_style2)
            #
            #             row_num += 1
            #         wb.save(response)
            #         return response
            #     except Exception as ex:
            #         pass

            elif action == 'rechazar':
                try:
                    data['title'] = u'Rechazar Solicitud'
                    data['solicitudes'] = SolicitudPublicacion.objects.get(pk=request.GET['id'])
                    return render(request, "inv_libros/rechazar.html", data)
                except Exception as ex:
                    pass

            elif action == 'evidenciaslibrocodigo':
                try:
                    data['title'] = u'Descargar Evidencias por Código'
                    template = get_template("inv_libros/evidenciaslibroscodigo.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content, 'title': data['title']})
                except Exception as ex:
                    msg = ex.__str__()
                    return JsonResponse({"result": "bad", "titulo": "Error", "mensaje": u"Error al obtener los datos. [%s]" % msg, "showSwal": "True", "swalType": "error"})

            elif action == 'evidenciascapitulocodigo':
                try:
                    data['title'] = u'Descargar Evidencias por Código'
                    template = get_template("inv_libros/evidenciascapituloscodigo.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content, 'title': data['title']})
                except Exception as ex:
                    msg = ex.__str__()
                    return JsonResponse({"result": "bad", "titulo": "Error", "mensaje": u"Error al obtener los datos. [%s]" % msg, "showSwal": "True", "swalType": "error"})

            # elif action == 'descargarevidenciaslibros':
            #     try:
            #         url = '/media/zipav/librosevidencias.zip'
            #         fantasy_zip = zipfile.ZipFile(SITE_STORAGE + url, 'w')
            #
            #         libros = LibroInvestigacion.objects.filter(status=True, fechapublicacion__isnull=False).order_by('id')
            #         for libro in libros:
            #             libro_id = libro.id
            #             nombre = "LIBRO_" + str(libro_id).zfill(4)
            #
            #             for evidencia in libro.detalleevidencias_set.filter(status=True):
            #                 ext = evidencia.archivo.__str__()[evidencia.archivo.__str__().rfind("."):]
            #                 if evidencia.evidencia.id == 7:
            #                     if os.path.exists(SITE_STORAGE + evidencia.archivo.url):
            #                         fantasy_zip.write(SITE_STORAGE + evidencia.archivo.url, nombre + '_ARCHIVODIGITALLIBRO_' + str(libro.fechapublicacion.year) + ext.lower())
            #
            #         fantasy_zip.close()
            #
            #         response = HttpResponse(open(SITE_STORAGE + url, 'rb'), content_type='application/zip')
            #         response['Content-Disposition'] = 'attachment; filename=librosevidencias' + random.randint(1, 10000).__str__() + '.zip'
            #         return response
            #     except Exception as ex:
            #         pass

            # elif action == 'descargarevidenciascapituloslibros':
            #     try:
            #         url = '/media/zipav/capituloslibrosevidencias.zip'
            #         fantasy_zip = zipfile.ZipFile(SITE_STORAGE + url, 'w')
            #
            #         capitulos = CapituloLibroInvestigacion.objects.filter(status=True, fechapublicacion__isnull=False).order_by('id')
            #         for capitulo in capitulos:
            #             capitulo_id = capitulo.id
            #             nombre = "CAPITULO_LIBRO_" + str(capitulo_id).zfill(4)
            #
            #             for evidencia in capitulo.detalleevidencias_set.filter(status=True):
            #                 ext = evidencia.archivo.__str__()[evidencia.archivo.__str__().rfind("."):]
            #                 if evidencia.evidencia.id == 11:
            #                     if os.path.exists(SITE_STORAGE + evidencia.archivo.url):
            #                         fantasy_zip.write(SITE_STORAGE + evidencia.archivo.url, nombre + '_LIBRO_' + str(capitulo.fechapublicacion.year) + ext.lower())
            #
            #                 elif evidencia.evidencia.id == 12:
            #                     if os.path.exists(SITE_STORAGE + evidencia.archivo.url):
            #                         fantasy_zip.write(SITE_STORAGE + evidencia.archivo.url, nombre + '_CAPITULODELIBRO_' + str(capitulo.fechapublicacion.year) + ext.lower())
            #
            #                 elif evidencia.evidencia.id == 14:
            #                     if os.path.exists(SITE_STORAGE + evidencia.archivo.url):
            #                         fantasy_zip.write(SITE_STORAGE + evidencia.archivo.url, nombre + '_FICHACATALOGRAFICA_' + str(capitulo.fechapublicacion.year) + ext.lower())
            #
            #                 elif evidencia.evidencia.id == 13:
            #                     if os.path.exists(SITE_STORAGE + evidencia.archivo.url):
            #                         fantasy_zip.write(SITE_STORAGE + evidencia.archivo.url, nombre + '_INFORMEREVISIONPARES_' + str(capitulo.fechapublicacion.year) + ext.lower())
            #
            #         fantasy_zip.close()
            #
            #         response = HttpResponse(open(SITE_STORAGE + url, 'rb'), content_type='application/zip')
            #         response['Content-Disposition'] = 'attachment; filename=capituloslibrosevidencias' + random.randint(1, 10000).__str__() + '.zip'
            #         return response
            #     except Exception as ex:
            #         pass

            return HttpResponseRedirect(request.path)

        else:
            data['title'] = u'Listado de Libros'
            search = None
            ids = None
            tipobus = None
            inscripcionid = None
            if 'id' in request.GET:
                data['tipobus'] = 2
                ids = request.GET['id']
                librosinvestigacion = LibroInvestigacion.objects.select_related().filter(Q(status=True) | Q(status=False, eliminadoxdoc=True), pk=int(encrypt(ids))).order_by('nombrelibro')
            elif 's' in request.GET:
                search = request.GET['s']
                if search.isdigit():
                    if int(request.GET['tipobus']) == 4:
                        data['tipobus'] = 4
                        anio = int(search)
                        librosinvestigacion = LibroInvestigacion.objects.select_related().filter(Q(status=True) | Q(status=False, eliminadoxdoc=True), fechapublicacion__year=anio)
                    else:
                        data['tipobus'] = 2
                        librosinvestigacion = LibroInvestigacion.objects.select_related().filter(Q(status=True) | Q(status=False, eliminadoxdoc=True), pk=search)
                else:
                    tipobus = int(request.GET['tipobus'])
                    data['tipobus'] = tipobus
                    if tipobus == 1:
                        librosinvestigacion = LibroInvestigacion.objects.select_related().filter(Q(status=True) | Q(status=False, eliminadoxdoc=True), revista__nombre__icontains=search, revista__status=True)
                    if tipobus == 2:
                        librosinvestigacion = LibroInvestigacion.objects.select_related().filter(Q(status=True) | Q(status=False, eliminadoxdoc=True), nombrelibro__icontains=search)
                    if tipobus == 3:
                        if ' ' in search:
                            s = search.split(" ")
                            participanteslibrosdoc = ParticipanteLibros.objects.values_list('libros_id').filter(status=True).filter(Q(profesor__persona__apellido1__contains=s[0]) & Q(profesor__persona__apellido2__contains=s[1])).distinct()
                            librosinvestigaciondoc = LibroInvestigacion.objects.select_related().filter(Q(status=True) | Q(status=False, eliminadoxdoc=True), pk__in=participanteslibrosdoc)

                            participanteslibrosadm = ParticipanteLibros.objects.values_list('libros_id').filter(status=True).filter(Q(administrativo__persona__apellido1__contains=s[0]) & Q(administrativo__persona__apellido2__contains=s[1])).distinct()
                            librosinvestigacionadm = LibroInvestigacion.objects.select_related().filter(Q(status=True) | Q(status=False, eliminadoxdoc=True), pk__in=participanteslibrosadm)
                            librosinvestigacion = librosinvestigaciondoc | librosinvestigacionadm
                        else:
                            participanteslibrosdoc = ParticipanteLibros.objects.values_list('libros_id').filter(status=True).filter(Q(profesor__persona__nombres__icontains=search) |
                                                                                                                                    Q(profesor__persona__apellido1__icontains=search) |
                                                                                                                                    Q(profesor__persona__apellido2__icontains=search)).distinct()
                            librosinvestigaciondoc = LibroInvestigacion.objects.select_related().filter(Q(status=True) | Q(status=False, eliminadoxdoc=True), pk__in=participanteslibrosdoc)

                            participanteslibrosadm = ParticipanteLibros.objects.values_list('libros_id').filter(status=True).filter(Q(administrativo__persona__nombres__icontains=search) |
                                                                                                                                    Q(administrativo__persona__apellido1__icontains=search) |
                                                                                                                                    Q(administrativo__persona__apellido2__icontains=search)).distinct()
                            librosinvestigacionadm = LibroInvestigacion.objects.select_related().filter(Q(status=True) | Q(status=False, eliminadoxdoc=True), pk__in=participanteslibrosadm)
                            librosinvestigacion = librosinvestigaciondoc | librosinvestigacionadm

            elif 'idsp' in request.GET:
                data['tipobus'] = 3
                lista_ids_personas = request.GET['idsp'].split(",")

                participanteslibros = ParticipanteLibros.objects.values_list('libros_id').filter(status=True).filter(Q(profesor__persona__id__in=lista_ids_personas) |
                                                                                                                        Q(administrativo__persona__id__in=lista_ids_personas)
                                                                                                                        ).distinct()
                librosinvestigacion = LibroInvestigacion.objects.select_related().filter(Q(status=True) | Q(status=False, eliminadoxdoc=True), pk__in=participanteslibros)
            else:
                data['tipobus'] = 2
                librosinvestigacion = LibroInvestigacion.objects.select_related().filter(Q(status=True) | Q(status=False, eliminadoxdoc=True)).order_by('nombrelibro')

            librosinvestigacion = librosinvestigacion.order_by('-fechapublicacion')

            estadolibro = 0
            if 'estadolibro' in request.GET:
                estadolibro = int(request.GET['estadolibro'])
                if estadolibro > 0:
                    if estadolibro == 1:
                        librosinvestigacion = librosinvestigacion.filter(aprobado=True)
                    else:
                        librosinvestigacion = librosinvestigacion.filter(aprobado=False)

            paging = MiPaginador(librosinvestigacion, 25)
            p = 1
            try:
                paginasesion = 1
                if 'paginador' in request.session:
                    paginasesion = int(request.session['paginador'])
                if 'page' in request.GET:
                    p = int(request.GET['page'])
                else:
                    p = paginasesion
                try:
                    page = paging.page(p)
                except:
                    p = 1
                page = paging.page(p)
            except:
                page = paging.page(p)
            request.session['paginador'] = p
            data['paging'] = paging
            data['page'] = page
            data['rangospaging'] = paging.rangos_paginado(p)
            data['libros'] = page.object_list
            data['search'] = search if search else ""
            data['ids'] = ids if ids else ""
            data['totallibros'] = total = librosinvestigacion.count()
            data['totalaprobados'] = totalaprobados = librosinvestigacion.filter(aprobado=True).count()
            data['totalporaprobar'] = total - totalaprobados
            data['estadolibro'] = estadolibro
            data['aniosevidencias'] = LibroInvestigacion.objects.filter(Q(status=True) | Q(status=False, eliminadoxdoc=True), fechapublicacion__isnull=False).annotate(anio=ExtractYear('fechapublicacion')).values_list('anio', flat=True).order_by('-anio').distinct()
            data['participantes'] = Persona.objects.values('id', 'cedula', 'nombres', 'apellido1', 'apellido2').filter(Q(profesor__participantelibros__isnull=False, profesor__participantelibros__status=True))

            return render(request, "inv_libros/view.html", data)
