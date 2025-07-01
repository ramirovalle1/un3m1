import json
import random
import sys
import time

import openpyxl
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.template.loader import get_template
from django.forms import model_to_dict
from decorators import last_access, secure_module
from django.template import Context
from django.db import transaction
from django.db.models import Max
from django.http import HttpResponseRedirect, JsonResponse, HttpResponse
from django.shortcuts import render
from django.db.models.query_utils import Q
from datetime import datetime

from postulate.commonviews import actualizar_nota_postulante
from postulate.models import Convocatoria, Partida, ConvocatoriaTerminosCondiciones, PartidaAsignaturas, \
    PersonaAplicarPartida, PersonaIdiomaPartida, PersonaFormacionAcademicoPartida, PersonaExperienciaPartida, \
    PersonaCapacitacionesPartida, PersonaPublicacionesPartida, CalificacionPostulacion, PartidaTribunal, \
    PersonaApelacion, ModeloEvaluativoDisertacion, CalificacionDisertacion, DetalleCalificacionDisertacion, \
    CalificacionEntrevista, DetalleCalificacionEntrevista, DetalleModeloEvaluativoConvocatoria, \
    ModeloEvaluativoConvocatoria, NotificacionGanador
from postulate.postular import validar_campos
from sagest.models import Departamento
from settings import EMAIL_INSTITUCIONAL_AUTOMATICO, ACTUALIZAR_FOTO_ALUMNOS
from sga.commonviews import adduserdata
from postulate.forms import ConvocatoriaForm, PartidaForm, ConvocatoriaTerminosForm, SubirArchivoForm, \
    SubirArchivoTestPsicologicoForm
from sga.funciones import log, MiPaginador, null_to_decimal, generar_nombre, logobjeto
from sga.funcionesxhtml2pdf import conviert_html_to_pdf
from sga.models import AreaConocimientoTitulacion, SubAreaConocimientoTitulacion, \
    SubAreaEspecificaConocimientoTitulacion, Carrera, Asignatura, Titulo, miinstitucion, CUENTAS_CORREOS, Persona
from sga.tasks import send_html_mail
from sga.templatetags.sga_extras import encrypt

from xlwt import *
import xlwt

import xlsxwriter
import io

@login_required(redirect_field_name='ret', login_url='/loginpostulate')
@secure_module
@last_access
# @transaction.atomic()
def view(request):
    global ex
    data = {}
    perfilprincipal, persona, periodo = request.session['perfilprincipal'], request.session['persona'], request.session['periodo']
    data['hoy'] = hoy = datetime.now().date()
    data['currenttime'] = datetime.now()
    data['perfil'] = persona.mi_perfil()
    data['periodo'] = periodo
    data['url_'] = request.path

    if request.method == 'POST':
        action = request.POST['action']

        if action == 'iniciardisertacion':
            try:
                with transaction.atomic():
                    postulante = PersonaAplicarPartida.objects.get(id=int(request.POST['id']))
                    modeloevaluativo = ModeloEvaluativoDisertacion.objects.filter(status=True, vigente=True,pk =postulante.partida.convocatoria.modeloevaluativo.id).first()
                    cal_ = CalificacionDisertacion(postulacion=postulante, modeloevaluativo=modeloevaluativo)
                    cal_.save(request)
                    for mod_ in modeloevaluativo.traer_aspectos():
                        det = DetalleCalificacionDisertacion(calificacion=cal_, parametro=mod_)
                        det.save(request)
                    log(u'Iniciar Disertación: %s' % cal_, request, "add")
                    res_json = {"error": False}
            except Exception as ex:
                res_json = {'error': True, "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)

        if action == 'iniciarentrevista':
            try:
                with transaction.atomic():
                    postulante = PersonaAplicarPartida.objects.get(id=int(request.POST['id']))
                    for cal in CalificacionEntrevista.objects.filter(postulacion=postulante, status=True):
                        cal.status = False
                        cal.save(request)
                    cal_ = CalificacionEntrevista(postulacion=postulante)
                    cal_.save(request)
                    for trib_ in PartidaTribunal.objects.filter(status=True, partida_id=postulante.partida.id, tipo=2):
                        det = DetalleCalificacionEntrevista(calificacion=cal_, tribunal=trib_)
                        det.save(request)
                    log(u'Iniciar Entrevista: %s' % cal_, request, "add")
                    res_json = {"error": False}
            except Exception as ex:
                res_json = {'error': True, "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)

        if action == 'calificarpunto':
            try:
                instance = DetalleCalificacionDisertacion.objects.get(id=int(request.POST['id']))
                instance.valor = request.POST['value']
                instance.save(request)
                log(u'{} : Calificación Nota Disertación Guardada'.format(instance.__str__()), request, "add")
                return JsonResponse({"result": "ok", "porcentaje": instance.get_porcentaje(), "puntos": instance.get_puntos(), "totalpuntos": instance.calificacion.totalpuntos(), "totalporcentaje": instance.calificacion.totalporcentaje()})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos. {}".format(str(ex))})

        if action == 'finalizardisertacion':
            try:
                with transaction.atomic():
                    filtro = CalificacionDisertacion.objects.get(id=int(request.POST['id']))
                    filtro.finalizada = True
                    filtro.observacion = request.POST['obs']
                    filtro.totalporcentaje = filtro.totalporcentaje()
                    filtro.notadisertacion = filtro.totalpuntos()
                    filtro.revisado_por = request.user
                    filtro.fecha_revision = datetime.now()
                    filtro.save(request)
                    for parametros in filtro.get_parametros():
                        parametros.porcentaje = parametros.get_porcentaje()
                        parametros.puntos = parametros.get_puntos()
                        parametros.save(request)
                    postulacion = PersonaAplicarPartida.objects.get(pk=filtro.postulacion.id)
                    postulacion.nota_final_disertacion = filtro.totalpuntos()
                    postulacion.save(request)
                    log(u'Finalizo Revisión de Disertación: %s' % filtro, request, "add")
                    res_json = {"error": False}
            except Exception as ex:
                res_json = {'error': True, "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)

        if action == 'calificarnotaentrevista':
            try:
                instance = DetalleCalificacionEntrevista.objects.get(id=int(request.POST['id']))
                instance.nota = request.POST['value']
                instance.save(request)
                log(u'{} : Calificación Nota Entrevista Guardada'.format(instance.__str__()), request, "add")
                return JsonResponse({"result": "ok", "notapromedio": instance.calificacion.nota_promedio(), "total_notas": instance.calificacion.total_notas()})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos. {}".format(str(ex))})

        if action == 'obstribunalentrevista':
            try:
                instance = DetalleCalificacionEntrevista.objects.get(id=int(request.POST['id']))
                instance.observacion = request.POST['value']
                instance.save(request)
                log(u'{} : Observación Tribunal Entrevista'.format(instance.__str__()), request, "add")
                return JsonResponse({"result": "ok"})
            except Exception as ex:

                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos. {}".format(str(ex))})

        if action == 'observacion':
            try:
                instance = PersonaAplicarPartida.objects.get(id=int(request.POST['id']))
                instance.observacion = request.POST['value']
                instance.save(request)
                log(u'{} : Observación Tribunal Entrevista'.format(instance.__str__()), request, "add")
                return JsonResponse({"result": "ok"})
            except Exception as ex:

                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos. {}".format(str(ex))})

        if action == 'finalizarentrevista':
            try:
                with transaction.atomic():
                    filtro = CalificacionEntrevista.objects.get(id=int(request.POST['id']))
                    filtro.finalizada = True
                    filtro.notaentrevista = filtro.nota_promedio()
                    filtro.revisado_por = request.user
                    filtro.fecha_revision = datetime.now()
                    filtro.save(request)
                    postulacion = PersonaAplicarPartida.objects.get(pk=filtro.postulacion.id)
                    postulacion.nota_final_entrevista = filtro.nota_promedio()
                    postulacion.save(request)
                    log(u'Finalizo Revisión de Entrevista: %s' % filtro, request, "add")
                    res_json = {"error": False}
            except Exception as ex:
                res_json = {'error': True, "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)

        if action == 'finalizarsegundaetapa':
            try:
                with transaction.atomic():
                    filtro = PersonaAplicarPartida.objects.get(id=int(request.POST['id']))
                    if not filtro.partida.convocatoria.modeloevaluativoconvocatoria:
                        if not filtro.traer_calificacion_disertacion():
                            return JsonResponse({'error': True, "message": "No tiene nota de disertación."}, safe=False)
                        else:
                            if not filtro.traer_calificacion_disertacion().finalizada:
                                return JsonResponse({'error': True, "message": "No tiene nota de disertación."}, safe=False)
                        if not filtro.traer_calificacion_entrevista():
                            return JsonResponse({'error': True, "message": "No tiene nota de entrevista."}, safe=False)
                        else:
                            if not filtro.traer_calificacion_entrevista().finalizada:
                                return JsonResponse({'error': True, "message": "No tiene nota de disertación."}, safe=False)

                        filtro.nota_final_disertacion = filtro.traer_calificacion_disertacion().notadisertacion
                        filtro.nota_final_entrevista = filtro.traer_calificacion_entrevista().notaentrevista
                        filtro.nota_final = filtro.total_segunda_etapa()
                    filtro.finsegundaetapa = True
                    filtro.setapa_revisado_por = request.user
                    filtro.setapa_fecha_revision = datetime.now()
                    filtro.save(request)
                    filtro.actualiza_estadofinal()
                    log(u'Finalizo Revisión de Segunda Etapa: %s' % filtro, request, "add")
                    res_json = {"error": False}

            except Exception as ex:
                res_json = {'error': True, "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)

        if action == 'finalizarevisionp':
            try:
                with transaction.atomic():
                    filtro = Partida.objects.get(id=int(request.POST['id']))
                    idpostulantes = filtro.personaaplicarpartida_set.values_list('id',flat=True).filter(partida=filtro, status=True, esmejorpuntuado=True).order_by('-nota_final_meritos')
                    postulantes = PersonaAplicarPartida.objects.filter(id__in=idpostulantes).order_by('-nota_final_meritos')
                    estado=True
                    if filtro.cerrada:
                        estado=False
                    postulantes.update(finsegundaetapa=estado,setapa_fecha_revision=datetime.now())
                    filtro.cerrada=estado
                    filtro.save(request,update_fields=['cerrada'])
                    for postulante in postulantes:
                        postulante.actualiza_estadofinal()
                    log(u'Finalizo Revisión de Segunda Etapa: %s' % filtro, request, "add")
                    res_json = {"error": False}

            except Exception as ex:
                res_json = {'error': True, "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)

        if action == 'establecerganador':
            try:
                if PersonaAplicarPartida.objects.filter(status=True,partida_id = encrypt(request.POST['idpar']), esganador=True,finsegundaetapa=True).order_by('-nota_final').exists():
                    for per in PersonaAplicarPartida.objects.filter(status=True, partida_id = encrypt(request.POST['idpar']), esganador=True,finsegundaetapa=True):
                        per.esganador=False
                        per.save(request)
                per = PersonaAplicarPartida.objects.get(status=True,id=encrypt(request.POST['id']))
                per.esganador = request.POST['ganador']
                per.save(request)
                log('Se cambio el ganador %s - %s'%(per.persona, per.partida),request,'EditGanador')
                return JsonResponse({'result':True,'mensaje':'Ganador declarado correctamente!','idc':encrypt(per.id),'est':per.esganador})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': u'Error al guardar los datos'})

        if action == 'habilitarrevision':
            try:
                if PersonaAplicarPartida.objects.values('id').filter(status=True, pk = request.POST['id']).exists():
                    persona = PersonaAplicarPartida.objects.get(status=True, pk = request.POST['id'])
                    persona.finsegundaetapa=False
                    persona.save(request)
                    log('Se habilito la revision de la segunda estapa %s - %s '%(persona.persona, persona.partida),request,'EditHabilitar')
                    return JsonResponse({'result':True,'mensaje':'Ganador declarado correctamente!'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': u'Error al guardar los datos'})

        if action == 'savearchivo':
            try:
                persona = PersonaAplicarPartida.objects.get(pk=request.POST['id'])
                if 'archivo' in request.FILES:
                    newfile = request.FILES['archivo']
                    modelo = persona.calificaciondisertacion_set.filter(status=True)[0]
                    calificaciondisertacion = CalificacionDisertacion.objects.filter(status=True,postulacion=persona,modeloevaluativo=modelo.modeloevaluativo)[0]
                    if calificaciondisertacion.archivo:
                        log('Se edito el archivo de criterio de calificacion de la disertación %s los archivos %s -%s'%(calificaciondisertacion,calificaciondisertacion.archivo,newfile),request,'EditArchivo')
                    else:
                        log('Se agrego el archivo de criterio de calificacion de la disertación %s el archivo: %s' % (calificaciondisertacion,newfile), request, 'add')
                    calificaciondisertacion.archivo = newfile
                    calificaciondisertacion.save(request)
                    return JsonResponse({'result':'ok','mensaje':u'Archivo guardado correctamente'})
                else:
                    return JsonResponse({'result':'bad','mensaje':u'No hay archivo cargado'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', 'mensaje': u'Error al guardar los datos'})

        elif action == 'calificar':
            try:
                result = actualizar_nota_postulante(request)
                return JsonResponse(result)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'importarnotas':
            with transaction.atomic():
                try:
                    form = SubirArchivoForm(request.POST,request.FILES)
                    if not form.is_valid():
                        raise NameError(f"{[{k:v[0]} for k,v in form.errors.items()]}")
                    id = encrypt(request.POST['id'])
                    convocatoria = Convocatoria.objects.get(status=True,id=id)
                    excel = request.FILES['archivo']
                    campo = form.cleaned_data['campo']
                    wb = openpyxl.load_workbook(excel)
                    worksheet = wb.worksheets[0]
                    numlist = worksheet.rows
                    count = 0
                    nombre_campo = campo.nombre if campo else 'PT'
                    lista_error = []
                    for pst in numlist:
                        if count>0:
                            cedula_ = str(pst[4].value) if len(str(pst[4].value)) == 10 else '0' + str(pst[4].value)
                            notas_ = pst[7].value if pst[7].value else 0
                            if Persona.objects.filter(status=True,cedula=cedula_):
                                pers = Persona.objects.get(status=True, cedula=cedula_)
                                if PersonaAplicarPartida.objects.values('id').filter(status=True, persona=pers, partida__convocatoria=convocatoria).exists():
                                    aplica_partida = PersonaAplicarPartida.objects.get(status=True, persona=pers, partida__convocatoria=convocatoria)
                                    partida = aplica_partida.partida
                                    modeloevaluativo = partida.convocatoria.modeloevaluativoconvocatoria
                                    if not modeloevaluativo:
                                        lista_error.append(f"Convocatoria sin modelo evaluativo: {partida.convocatoria}, persona ci: {cedula_}")
                                    try:
                                        aplica_partida.valor_nombre_campo(nombre_campo)
                                    except:
                                        lista_error.append(f'Error al crear modelo evaluativo Persona: {cedula_}. Nota: {notas_}')
                                        print(f'Error al crear modelo evaluativo Persona: {cedula_}. Nota: {notas_}')
                                        continue
                                    try:
                                        actualizar_nota_postulante_2(request, aplica_partida, nombre_campo, notas_)
                                    except:
                                        lista_error.append(f'Error al guardar la nota Persona: {cedula_}. Nota: {notas_}')
                                        print(f'Error al guardar la nota Persona: {cedula_}. Nota: {notas_}')
                                        continue
                                else:
                                    lista_error.append(f'No se encontro postulación de la Persona: {cedula_}. Nota: {notas_}')
                            else:
                                lista_error.append(f'No se encontro postulación de la Persona: {cedula_}. Nota: {notas_}')
                        count +=1
                    if lista_error:
                        res_js = {'result':False,"mensaje":f"Hay conflictos: {lista_error}",'modalsuccess':True}
                    else:
                        res_js = {'result':False}
                except Exception as ex:
                    transaction.set_rollback(True)
                    err_ = f"Ocurrio un error: {ex.__str__()}. En la linea {sys.exc_info()[-1].tb_lineno}"
                    res_js = {'result':True,'mensaje':err_}
                return JsonResponse(res_js)

        elif action == 'loadtestpsc':
            try:
                idpap = encrypt(request.POST.get('idpersonapartida'))
                personaaplica = PersonaAplicarPartida.objects.get(status=True,id=int(idpap))
                form = SubirArchivoTestPsicologicoForm(request.POST,request.FILES)
                if not form.is_valid():
                    raise NameError(f'{[{k:v[0]} for k,v in form.errors.items()]}')
                if not 'archivopsc' in request.FILES:
                    raise NameError(f'Debe seleccionar un archivo')
                archivo = request.FILES['archivopsc']
                ext_ = archivo._name.split('.')[-1]
                if not ext_.lower() in ['pdf']:
                    raise NameError("Formato de archivo incorrecto, intente con el formato .pdf")
                if archivo.size > 30943040:
                    raise NameError("Archivo excede los 30MB")
                archivo._name = generar_nombre("pps_", archivo._name)
                personaaplica.archivopsc = archivo
                personaaplica.save(request,update_fields=['archivopsc'])
                res_js = {'result': False, 'id_partida':personaaplica.partida.id}
            except Exception as ex:
                transaction.set_rollback(True)
                err_ = f"Ocurrio un error: {ex}. En la linea {sys.exc_info()[-1].tb_lineno}"
                res_js = {'result':True,'mensaje':err_}
            return JsonResponse(res_js)

        elif action == 'notificacion_general':
            try:
                msg_convocatoria = request.POST.get('id_convocatoria_mensaje', None)
                msg_contenido = request.POST.get('mensajecontenido', None)
                msg_title = request.POST.get('titulomensaje', None)
                msg_contenido_header = request.POST.get('mensajecabecera', None)
                option = request.POST.get('envio_opciones', None)
                option = int(option) if option else None

                if option == 1:
                    envio_email_pt_nopasaron(request=request, msg_convocatoria=msg_convocatoria, msg_contenido=msg_contenido,
                                             msg_title=msg_title, msg_contenido_header=msg_contenido_header)


                res_js = {'result':False}
            except Exception as ex:
                err_ = f'Se ha presentado el siguiente problema: {ex.__str__()}({sys.exc_info()[-1].tb_lineno})'
                res_js = {'result':True, 'mensaje':err_}
            return JsonResponse(res_js)

        elif action == 'loadnotificacionentrevista':
            try:
                msg_contenido = request.POST.get('mensajecontenido', None)
                msg_title = request.POST.get('titulomensaje', None)
                msg_contenido_header = request.POST.get('mensajecabecera', None)
                id = request.POST.get('id', None)
                peraplica = PersonaAplicarPartida.objects.get(status=True,id=int(encrypt(id)))
                if peraplica.persona:
                    send_html_mail(f"Postulate: {msg_title}",
                                   "emails/segundaetapa_postulate.html",
                                   {'sistema': u'SISTEMA POSTULATE UNEMI', 'persona': peraplica.persona,
                                    'partida': peraplica.partida, 'contenido': msg_contenido,
                                    'contenido_static': msg_contenido_header,
                                    'title': msg_title,
                                    't': miinstitucion()}, [peraplica.persona.email], [],
                                   cuenta=CUENTAS_CORREOS[30][1])
                res_js = {'result':False, 'id_partida':peraplica.partida.id}
            except Exception as ex:
                err_ = f'Se ha presentado el siguiente problema: {ex.__str__()}({sys.exc_info()[-1].tb_lineno})'
                res_js = {'result': True, 'mensaje': err_}
            return JsonResponse(res_js)

        elif action == 'loadnotificacionentr3pt':
            try:
                msg_contenido = request.POST.get('mensajecontenido', None)
                msg_title = request.POST.get('titulomensaje', None)
                msg_contenido_header = request.POST.get('mensajecabecera', None)
                id = request.POST.get('id', None)
                partida = Partida.objects.get(status=True,id=int(encrypt(id)))
                envio_email_3mejores_pt(request=request, partida=partida,
                                         msg_contenido=msg_contenido,
                                         msg_title=msg_title, msg_contenido_header=msg_contenido_header)
                res_js = {'result':False}
            except Exception as ex:
                err_ = f'Se ha presentado el siguiente problema: {ex.__str__()}({sys.exc_info()[-1].tb_lineno})'
                res_js = {'result': True, 'mensaje': err_}
            return JsonResponse(res_js)

        elif action == 'loadnotificacionentrexcl3pt':
            try:
                msg_contenido = request.POST.get('mensajecontenido', None)
                msg_title = request.POST.get('titulomensaje', None)
                msg_contenido_header = request.POST.get('mensajecabecera', None)
                id = request.POST.get('id', None)
                partida = Partida.objects.get(status=True,id=int(encrypt(id)))
                envio_email_excl3mejores_pt(request=request, partida=partida,
                                         msg_contenido=msg_contenido,
                                         msg_title=msg_title, msg_contenido_header=msg_contenido_header)
                res_js = {'result':False}
            except Exception as ex:
                err_ = f'Se ha presentado el siguiente problema: {ex.__str__()}({sys.exc_info()[-1].tb_lineno})'
                res_js = {'result': True, 'mensaje': err_}
            return JsonResponse(res_js)

        elif action == 'loadnotificaciomenor70pt':
            try:
                msg_contenido = request.POST.get('mensajecontenido', None)
                msg_title = request.POST.get('titulomensaje', None)
                msg_contenido_header = request.POST.get('mensajecabecera', None)
                id = request.POST.get('id', None)
                partida = Partida.objects.get(status=True,id=int(encrypt(id)))
                envio_email_m70_pt(request=request, partida=partida,
                                         msg_contenido=msg_contenido,
                                         msg_title=msg_title, msg_contenido_header=msg_contenido_header)
                res_js = {'result':False}
            except Exception as ex:
                err_ = f'Se ha presentado el siguiente problema: {ex.__str__()}({sys.exc_info()[-1].tb_lineno})'
                res_js = {'result': True, 'mensaje': err_}
            return JsonResponse(res_js)

        elif action == 'loadnotificacionotafinalupper':
            try:
                msg_contenido = request.POST.get('mensajecontenido', None)
                msg_title = request.POST.get('titulomensaje', None)
                msg_contenido_header = request.POST.get('mensajecabecera', None)
                id = request.POST.get('id', None)
                partida = Partida.objects.get(status=True,id=int(encrypt(id)))
                envio_email_notafina_70up(request=request, partida=partida,
                                         msg_contenido=msg_contenido,
                                         msg_title=msg_title, msg_contenido_header=msg_contenido_header)
                res_js = {'result':False}
            except Exception as ex:
                err_ = f'Se ha presentado el siguiente problema: {ex.__str__()}({sys.exc_info()[-1].tb_lineno})'
                res_js = {'result': True, 'mensaje': err_}
            return JsonResponse(res_js)

        elif action == 'loadnotificacionotafinallower':
            try:
                msg_contenido = request.POST.get('mensajecontenido', None)
                msg_title = request.POST.get('titulomensaje', None)
                msg_contenido_header = request.POST.get('mensajecabecera', None)
                id = request.POST.get('id', None)
                partida = Partida.objects.get(status=True,id=int(encrypt(id)))
                envio_email_notafina_70down(request=request, partida=partida,
                                         msg_contenido=msg_contenido,
                                         msg_title=msg_title, msg_contenido_header=msg_contenido_header)
                res_js = {'result':False}
            except Exception as ex:
                err_ = f'Se ha presentado el siguiente problema: {ex.__str__()}({sys.exc_info()[-1].tb_lineno})'
                res_js = {'result': True, 'mensaje': err_}
            return JsonResponse(res_js)


        return JsonResponse({"result": "bad", "mensaje": u"Solicitud Incorrecta."})
    else:
        adduserdata(request, data)
        if 'action' in request.GET:
            data['action'] = action = request.GET['action']

            if action == 'buscarpartidas':
                try:
                    convocatoria = Convocatoria.objects.get(pk=request.GET['convocatoria'])
                    qspartidas = Partida.objects.filter(convocatoria=convocatoria, status=True)
                    validarpartidas = True
                    departamentogestion = Departamento.objects.filter(status=True, permisodepartamento=2)
                    if departamentogestion.exists():
                        if departamentogestion.first().mis_integrantes().filter(id=persona.pk).exists():
                            validarpartidas = False
                    if request.user.is_superuser:
                        validarpartidas = False
                    if validarpartidas:
                        idsmispartidas = PartidaTribunal.objects.filter(status=True, persona=persona, tipo=2).values_list('partida_id', flat=True)
                        qspartidas = qspartidas.filter(id__in=idsmispartidas)
                    data['partidas'] = qspartidas.order_by('codpartida')
                    template = get_template("postulate/adm_segundaetapa/cbpartidas.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            if action == 'buscarpostulantes':
                try:
                    data['partida'] = partida = Partida.objects.get(pk=request.GET['partida'])
                    data['tribunal'] = PartidaTribunal.objects.filter(status=True, partida=partida, tipo=2)
                    nummejores = partida.convocatoria.nummejorespuntuados
                    data['postulantes'] = PersonaAplicarPartida.objects.filter(partida=partida, status=True,esmejorpuntuado=True, estado__in=[1,4,5]).order_by('-nota_final_meritos')
                    template = get_template("postulate/adm_segundaetapa/postulantes.html")
                    if partida.convocatoria.modeloevaluativoconvocatoria:
                        data['modelo'] = partida.convocatoria.modeloevaluativoconvocatoria
                        template = get_template("postulate/adm_segundaetapa/postulantesv2.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            if action == 'evaluardisertacion':
                try:
                    data['id'] = id = int(request.GET['id'])
                    data['postulante'] = postulante = PersonaAplicarPartida.objects.get(pk=id)
                    data['partida'] = partida = Partida.objects.get(pk=postulante.partida.pk)
                    data['modeloevaluativo'] = modeloevaluativo = ModeloEvaluativoDisertacion.objects.filter(status=True, vigente=True,pk =partida.convocatoria.modeloevaluativo.id).first()
                    data['calificacion'] = calificacion = CalificacionDisertacion.objects.filter(postulacion=postulante, status=True).first()
                    template = get_template("postulate/adm_segundaetapa/calificardisertacion.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            if action == 'evaluarentrevista':
                try:
                    data['id'] = id = int(request.GET['id'])
                    data['postulante'] = postulante = PersonaAplicarPartida.objects.get(pk=id)
                    data['partida'] = partida = Partida.objects.get(pk=postulante.partida.pk)
                    data['tribunal'] = tribunal = PartidaTribunal.objects.filter(status=True, partida=partida, tipo=2).order_by('cargos')
                    data['calificacion'] = calificacion = CalificacionEntrevista.objects.filter(postulacion=postulante, status=True).first()
                    template = get_template("postulate/adm_segundaetapa/calificarentrevista.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, 'mensaje': str(ex)})

            if action == 'calificar':
                try:
                    data['id'] = id = int(request.GET['id'])
                    data['postulante'] = postulante = PersonaAplicarPartida.objects.get(pk=id)
                    data['partida'] = partida = Partida.objects.get(pk=postulante.partida.pk)
                    data['modeloevaluativo'] = partida.convocatoria.modeloevaluativoconvocatoria
                    data['tribunal'] = tribunal = PartidaTribunal.objects.filter(status=True, partida=partida, tipo=2).order_by('cargos')
                    template = get_template("postulate/adm_segundaetapa/calificar.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, 'mensaje': str(ex)})

            if action == 'excel_postulantes__all':
                try:
                    output = io.BytesIO()
                    workbook = xlsxwriter.Workbook(output)
                    qspartidas = Partida.objects.filter(status=True, convocatoria__vigente=True)
                    partidasmodelo = qspartidas.filter(convocatoria__modeloevaluativoconvocatoria__isnull=False)
                    modelos = ModeloEvaluativoConvocatoria.objects.filter(status=True, id__in=(partidasmodelo.values_list('convocatoria__modeloevaluativoconvocatoria_id')))
                    partidasnormal = qspartidas.filter(convocatoria__modeloevaluativoconvocatoria__isnull=True)
                    formato_titulo = workbook.add_format(
                        {'align': 'center', 'bold': 1, 'border': 1, 'text_wrap': True, 'fg_color': '#0b2f44',
                         'font_color': 'white', 'size': 16})
                    formato_cabecera = workbook.add_format(
                        {'align': 'center', 'bold': 1, 'border': 1, 'text_wrap': True, 'fg_color': '#0b2f44',
                         'font_color': 'white'})
                    style2 = workbook.add_format(
                        {'text_wrap': True, 'align': 'center', 'valign': 'vcenter', 'border': 1})

                    for modeloevaluativo in modelos:
                        ws = workbook.add_worksheet(str(modeloevaluativo.id))
                        ws.set_column(0, 28, 25)
                        ws.write( 2 ,0, 'Convocatoria', formato_cabecera)
                        ws.write( 2 ,1, 'Carrera', formato_cabecera)
                        ws.write( 2 ,2, 'Categoria', formato_cabecera)
                        ws.write( 2 ,3, 'Partida', formato_cabecera)
                        ws.write( 2 ,4, 'Dedicacion', formato_cabecera)
                        ws.write( 2 ,5,  'Cod. Unico', formato_cabecera)
                        ws.write( 2 ,6, 'Postulante', formato_cabecera)
                        ws.write( 2 ,7, 'Identificación', formato_cabecera)
                        ws.write( 2 ,8, 'Correo', formato_cabecera)
                        ws.write( 2 ,9, 'Telf', formato_cabecera)
                        ws.write( 2 ,10, 'Tema', formato_cabecera)
                        ws.write( 2 ,11, 'Fecha', formato_cabecera)
                        ws.write( 2 ,12, 'Hora', formato_cabecera)
                        ws.write( 2 ,13, 'Lugar', formato_cabecera)
                        ws.write( 2 ,14, 'Obs', formato_cabecera)
                        ws.write( 2 ,15, 'Obs. general', formato_cabecera)
                        ws.write( 2 ,16, 'Estado', formato_cabecera)
                        row_num = 3
                        col_num = 17
                        camposm = modeloevaluativo.campos()
                        for campo in camposm:
                            ws.write( 2,col_num, str(campo.nombre), formato_cabecera)
                            col_num+=1
                        ws.write( 2 ,col_num, 'Puntaje final', formato_cabecera)
                        ws.merge_range(0,0,0,col_num, 'UNIVERSIDAD ESTATAL ESTATAL DE MILAGRO', formato_titulo)
                        ws.merge_range(1,0,1,col_num, 'LISTADO DE POSTULANTES MEJORES PUNTUADOS', formato_titulo)
                        partidas = partidasmodelo.filter(convocatoria__modeloevaluativoconvocatoria=modeloevaluativo)
                        for partidas in partidas:
                            listado = partidas.participantes_mejores_puntuados()
                            for det in listado:
                                carrera=det.partida.carrera.__str__() if det.partida.carrera else ''
                                ws.write(row_num, 0, det.partida.convocatoria.descripcion, style2)
                                ws.write(row_num, 1, carrera, style2)
                                ws.write(row_num, 2, det.partida.convocatoria.tipocontrato.nombre if det.partida.convocatoria.tipocontrato else '', style2)
                                ws.write(row_num, 3, str(det.partida), style2)
                                ws.write(row_num, 4, det.partida.get_dedicacion_display(), style2)
                                ws.write(row_num, 5, det.pk, style2)
                                ws.write(row_num, 6, str(det.persona.nombre_completo_minus()), style2)
                                ws.write(row_num, 7, det.persona.cedula, style2)
                                ws.write(row_num, 8, det.persona.email, style2)
                                ws.write(row_num, 9, "{} {}".format(det.persona.telefono,det.persona.telefono_conv), style2)
                                tema, fechaasistencia, horaasistencia, lugar, observacion = '', '', '', '', ''
                                if det.traer_agenda_entrevista():
                                    tema = det.traer_agenda_entrevista().tema
                                    fechaasistencia = det.traer_agenda_entrevista().fechaasistencia
                                    horaasistencia = det.traer_agenda_entrevista().horasistencia
                                    lugar = det.traer_agenda_entrevista().lugar
                                    observacion = det.traer_agenda_entrevista().observacion
                                ws.write(row_num, 10, tema, style2)
                                ws.write(row_num, 11, str(fechaasistencia), style2)
                                ws.write(row_num, 12, str(horaasistencia), style2)
                                ws.write(row_num, 13, lugar, style2)
                                ws.write(row_num, 14, observacion, style2)
                                ws.write(row_num, 15, det.observacion, style2)
                                ws.write(row_num, 16, det.get_estado_display(), style2)
                                col_num = 17
                                for campo in camposm:
                                    resultado = det.campo(campo.nombre)
                                    ws.write(row_num, col_num, resultado.valor if resultado else '', style2)
                                    col_num+=1
                                ws.write(row_num, col_num, det.nota_final, style2)
                                row_num += 1
                    if partidasnormal:
                        ws = workbook.add_worksheet('partidas_sin_modelo')
                        ws.set_column(0, 28, 25)
                        ws.merge_range(0, 0, 0, 21, 'UNIVERSIDAD ESTATAL ESTATAL DE MILAGRO', formato_titulo)
                        ws.merge_range(1, 0, 1, 21, 'LISTADO DE POSTULANTES MEJORES PUNTUADOS', formato_titulo)
                        ws.write(2, 0, 'Convocatoria', formato_cabecera)
                        ws.write(2, 1, 'Categoria', formato_cabecera)
                        ws.write(2, 2, 'Partida', formato_cabecera)
                        ws.write(2, 3, 'Dedicacion', formato_cabecera)
                        ws.write(2, 4, 'Cod. Unico', formato_cabecera)
                        ws.write(2, 5, 'Apellidos', formato_cabecera)
                        ws.write(2, 6, 'Nombres', formato_cabecera)
                        ws.write(2, 7, 'Identificación', formato_cabecera)
                        ws.write(2, 8, 'Correo', formato_cabecera)
                        ws.write(2, 9, 'Telf', formato_cabecera)
                        ws.write(2, 10, 'Telf. Conv.', formato_cabecera)
                        ws.write(2, 11, 'Tema', formato_cabecera)
                        ws.write(2, 12, 'Fecha', formato_cabecera)
                        ws.write(2, 13, 'Hora', formato_cabecera)
                        ws.write(2, 14, 'Lugar', formato_cabecera)
                        ws.write(2, 15, 'Obs', formato_cabecera)
                        ws.write(2, 16, 'Nota Disertación.', formato_cabecera)
                        ws.write(2, 17, 'Nota Entrevista', formato_cabecera)
                        ws.write(2, 18, 'Nota Final', formato_cabecera)
                        ws.write(2, 19, 'Estado', formato_cabecera)
                        ws.write(2, 20, 'Fecha Revisión', formato_cabecera)
                        ws.write(2, 21, 'Revisado Por', formato_cabecera)
                        row_num = 3
                        for partidan in partidasnormal:
                            listado = partidan.participantes_mejores_puntuados()
                            for det in listado:
                                ws.write(row_num, 0, det.partida.convocatoria.descripcion, style2)
                                ws.write(row_num, 1,
                                         det.partida.convocatoria.tipocontrato.nombre if det.partida.convocatoria.tipocontrato else '',
                                         style2)
                                ws.write(row_num, 2, str(det.partida), style2)
                                ws.write(row_num, 3, det.partida.get_dedicacion_display(), style2)
                                ws.write(row_num, 4, det.pk, style2)
                                ws.write(row_num, 5, "{} {}".format(det.persona.apellido1, det.persona.apellido2),
                                         style2)
                                ws.write(row_num, 6, "{}".format(det.persona.nombres), style2)
                                ws.write(row_num, 7, det.persona.cedula, style2)
                                ws.write(row_num, 8, det.persona.email, style2)
                                ws.write(row_num, 9, det.persona.telefono, style2)
                                ws.write(row_num, 10, det.persona.telefono_conv, style2)
                                tema, fechaasistencia, horaasistencia, lugar, observacion = '', '', '', '', ''
                                if det.traer_agenda_entrevista():
                                    tema = det.traer_agenda_entrevista().tema
                                    fechaasistencia = det.traer_agenda_entrevista().fechaasistencia
                                    horaasistencia = det.traer_agenda_entrevista().horasistencia
                                    lugar = det.traer_agenda_entrevista().lugar
                                    observacion = det.traer_agenda_entrevista().observacion
                                ws.write(row_num, 11, tema, style2)
                                ws.write(row_num, 12, str(fechaasistencia), style2)
                                ws.write(row_num, 13, str(horaasistencia), style2)
                                ws.write(row_num, 14, lugar, style2)
                                ws.write(row_num, 15, observacion, style2)
                                nota70, nota30 = '', ''
                                if det.traer_calificacion_disertacion():
                                    if det.traer_calificacion_disertacion().finalizada:
                                        nota70 = det.traer_calificacion_disertacion().nota_porcentual_70()
                                if det.traer_calificacion_entrevista():
                                    if det.traer_calificacion_entrevista().finalizada:
                                        nota30 = det.traer_calificacion_entrevista().nota_porcentual_30()
                                ws.write(row_num, 16, nota70, style2)
                                ws.write(row_num, 17, nota30, style2)
                                if det.finsegundaetapa:
                                    ws.write(row_num, 18, det.total_segunda_etapa(), style2)
                                else:
                                    ws.write(row_num, 18, '', style2)
                                revisado_por, fecha_revision = '', ''
                                if det.finsegundaetapa:
                                    revisado_por = det.setapa_revisado_por.username if det.setapa_revisado_por else ''
                                    fecha_revision = str(det.setapa_fecha_revision)
                                ws.write(row_num, 19, '', style2)
                                ws.write(row_num, 20, revisado_por, style2)
                                ws.write(row_num, 21, fecha_revision, style2)
                                row_num += 1
                    workbook.close()
                    output.seek(0)
                    filename = 'total_Segunda_etapa.xlsx'
                    response = HttpResponse(output,
                                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    response['Content-Disposition'] = 'attachment; filename=%s' % filename
                    return response
                except Exception as ex:
                    messages.error(request, str(ex))

            if action == 'mejorpuntuadopartid':
                try:
                    __autor__ = 'unemi'
                    output = io.BytesIO()
                    workbook = xlsxwriter.Workbook(output)
                    ws = workbook.add_worksheet('mejor_puntuado')
                    title = workbook.add_format({'font_name': 'Calibri', 'align': 'center', 'bold': True, 'font_size': 18, 'valign': 'vcenter'})
                    title2 = workbook.add_format({'font_name': 'Calibri', 'align': 'center', 'bold': True, 'font_size': 14, 'valign': 'vcenter'})
                    fuentecabecera = workbook.add_format({'font_name': 'Calibri', 'border': 1, 'align': 'center', 'font_size': 11, 'valign': 'vcenter', 'bg_color': '#E4E5DF'})
                    style2 = workbook.add_format({'text_wrap': True, 'font_name': 'Calibri', 'border': 1, 'align': 'center', 'font_size': 11, 'valign': 'vcenter'})
                    ws.merge_range(0, 0, 0, 8, 'UNIVERSIDAD ESTATAL ESTATAL DE MILAGRO', title)
                    ws.merge_range(1, 1, 0, 8, 'LISTADO DE POSTULANTES MEJORES PUNTUADOS', title2)
                    row_num = 2
                    columns = [
                        ('Convocatoria', 90),
                        ('Categoria', 50),
                        ('Partida', 80),
                        ('Dedicacion', 15),
                        ('Cod. Unico', 15),
                        ('Apellidos', 30),
                        ('Nombres', 30),
                        ('Titulo', 50),
                        ('Archivo', 50),
                        ('Identificación', 20),
                        ('Correo', 30),
                        ('Telf.', 20),
                        ('Telf. Conv.', 20),
                        ('Tema', 100),
                        ('Fecha', 15),
                        ('Hora', 15),
                        ('Lugar', 25),
                        ('Obs', 60),
                        ('Nota Disertación.', 15),
                        ('Nota Entrevista', 15),
                        ('Nota Final', 15),
                        ('Revisado Por', 15),
                        ('Fecha Revisión', 40),
                    ]
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], fuentecabecera)
                        ws.set_column(col_num,col_num,columns[col_num][1])
                    row_num += 1
                    for partidas in Partida.objects.filter(status=True, convocatoria__vigente=True).order_by('convocatoria__descripcion'):
                        listado = partidas.personaaplicarpartida_set.filter(status=True,finsegundaetapa=True).order_by('-nota_final')[:1]
                        for det in listado:
                            if det.persona.mis_titulaciones().count() > 1:
                                row_count = row_num + det.persona.mis_titulaciones().count() - 1
                                ws.merge_range(row_num, 0, row_count, 0, det.partida.convocatoria.descripcion, style2)
                                ws.merge_range(row_num, 1, row_count, 1, det.partida.convocatoria.tipocontrato.nombre if det.partida.convocatoria.tipocontrato else '', style2)
                                ws.merge_range(row_num, 2, row_count, 2, str(det.partida), style2)
                                ws.merge_range(row_num, 3, row_count, 3, det.partida.get_dedicacion_display(), style2)
                                ws.merge_range(row_num, 4, row_count, 4, det.pk, style2)
                                ws.merge_range(row_num, 5, row_count, 5, "{} {}".format(det.persona.apellido1, det.persona.apellido2), style2)
                                ws.merge_range(row_num, 6, row_count, 6, "{}".format(det.persona.nombres), style2)
                                ws.merge_range(row_num, 9, row_count, 9, det.persona.cedula, style2)
                                ws.merge_range(row_num, 10, row_count, 10, det.persona.email, style2)
                                ws.merge_range(row_num, 11, row_count, 11, det.persona.telefono, style2)
                                ws.merge_range(row_num, 12, row_count, 12, det.persona.telefono_conv, style2)
                                tema, fechaasistencia, horaasistencia, lugar, observacion = '', '', '', '', ''
                                if det.traer_agenda_entrevista():
                                    tema = det.traer_agenda_entrevista().tema
                                    fechaasistencia = det.traer_agenda_entrevista().fechaasistencia
                                    horaasistencia = det.traer_agenda_entrevista().horasistencia
                                    lugar = det.traer_agenda_entrevista().lugar
                                    observacion = det.traer_agenda_entrevista().observacion
                                ws.merge_range(row_num, 13, row_count, 13, tema, style2)
                                ws.merge_range(row_num, 14, row_count, 14, str(fechaasistencia), style2)
                                ws.merge_range(row_num, 15, row_count, 15, str(horaasistencia), style2)
                                ws.merge_range(row_num, 16, row_count, 16, lugar, style2)
                                ws.merge_range(row_num, 17, row_count, 17, observacion, style2)
                                nota70, nota30 = '', ''
                                if det.traer_calificacion_disertacion():
                                    if det.traer_calificacion_disertacion().finalizada:
                                        nota70 = det.traer_calificacion_disertacion().nota_porcentual_70()
                                if det.traer_calificacion_entrevista():
                                    if det.traer_calificacion_entrevista().finalizada:
                                        nota30 = det.traer_calificacion_entrevista().nota_porcentual_30()
                                ws.merge_range(row_num, 18, row_count, 18, nota70, style2)
                                ws.merge_range(row_num, 19, row_count, 19, nota30, style2)
                                if det.finsegundaetapa:
                                    ws.merge_range(row_num, 20, row_count, 20, det.total_segunda_etapa(), style2)
                                else:
                                    ws.merge_range(row_num, 20, row_count, 20, '', style2)
                                revisado_por, fecha_revision = '', ''
                                if det.finsegundaetapa:
                                    revisado_por = det.setapa_revisado_por.username if det.setapa_revisado_por else ''
                                    fecha_revision = str(det.setapa_fecha_revision)
                                ws.merge_range(row_num, 21, row_count, 21, revisado_por, style2)
                                ws.merge_range(row_num, 22, row_count, 22, fecha_revision, style2)
                                for titulo in det.persona.mis_titulaciones():
                                    ws.write(row_num, 7, "{}".format(titulo.titulo), style2)
                                    ws.write(row_num, 8, "https://sga.unemi.edu.ec/{}".format(titulo.archivo.url) if titulo.archivo else '', style2)
                                    row_num += 1
                            else:
                                ws.write(row_num, 0, det.partida.convocatoria.descripcion, style2)
                                ws.write(row_num, 1, det.partida.convocatoria.tipocontrato.nombre if det.partida.convocatoria.tipocontrato else '', style2)
                                ws.write(row_num, 2, str(det.partida), style2)
                                ws.write(row_num, 3, det.partida.get_dedicacion_display(), style2)
                                ws.write(row_num, 4, det.pk, style2)
                                ws.write(row_num, 5, "{} {}".format(det.persona.apellido1, det.persona.apellido2), style2)
                                ws.write(row_num, 6, "{}".format(det.persona.nombres), style2)
                                ws.write(row_num, 7, "", style2)
                                ws.write(row_num, 8, "", style2)
                                ws.write(row_num, 9, det.persona.cedula, style2)
                                ws.write(row_num, 10, det.persona.email, style2)
                                ws.write(row_num, 11, det.persona.telefono, style2)
                                ws.write(row_num, 12, det.persona.telefono_conv, style2)
                                tema, fechaasistencia, horaasistencia, lugar, observacion = '', '', '', '', ''
                                if det.traer_agenda_entrevista():
                                    tema = det.traer_agenda_entrevista().tema
                                    fechaasistencia = det.traer_agenda_entrevista().fechaasistencia
                                    horaasistencia = det.traer_agenda_entrevista().horasistencia
                                    lugar = det.traer_agenda_entrevista().lugar
                                    observacion = det.traer_agenda_entrevista().observacion
                                ws.write(row_num, 13, tema, style2)
                                ws.write(row_num, 14, str(fechaasistencia), style2)
                                ws.write(row_num, 15, str(horaasistencia), style2)
                                ws.write(row_num, 16, lugar, style2)
                                ws.write(row_num, 17, observacion, style2)
                                nota70, nota30 = '', ''
                                if det.traer_calificacion_disertacion():
                                    if det.traer_calificacion_disertacion().finalizada:
                                        nota70 = det.traer_calificacion_disertacion().nota_porcentual_70()
                                if det.traer_calificacion_entrevista():
                                    if det.traer_calificacion_entrevista().finalizada:
                                        nota30 = det.traer_calificacion_entrevista().nota_porcentual_30()
                                ws.write(row_num, 18, nota70, style2)
                                ws.write(row_num, 19, nota30, style2)
                                if det.finsegundaetapa:
                                    ws.write(row_num, 20, det.total_segunda_etapa(), style2)
                                else:
                                    ws.write(row_num, 20, '', style2)
                                revisado_por, fecha_revision = '', ''
                                if det.finsegundaetapa:
                                    revisado_por = det.setapa_revisado_por.username if det.setapa_revisado_por else ''
                                    fecha_revision = str(det.setapa_fecha_revision)
                                ws.write(row_num, 21, revisado_por, style2)
                                ws.write(row_num, 22, fecha_revision, style2)
                                row_num += 1
                    workbook.close()
                    output.seek(0)
                    filename = 'reporte_mejores_puntuados_' + random.randint(1, 10000).__str__() + '.xlsx'
                    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    response['Content-Disposition'] = 'attachment; filename=%s' % filename
                    return response
                except Exception as ex:
                    return "{}?info=ErrorLinea{}Error".format(request.path, str(sys.exc_info()[-1].tb_lineno), str(ex))

            if action == 'bancodatospartidas':
                try:
                    __autor__ = 'unemi'
                    output = io.BytesIO()
                    workbook = xlsxwriter.Workbook(output)
                    ws = workbook.add_worksheet('partidas')
                    title = workbook.add_format({'font_name': 'Calibri', 'align': 'center', 'bold': True, 'font_size': 18, 'valign': 'vcenter'})
                    title2 = workbook.add_format({'font_name': 'Calibri', 'align': 'center', 'bold': True, 'font_size': 14, 'valign': 'vcenter'})
                    fuentecabecera = workbook.add_format({'font_name': 'Calibri', 'border': 1, 'align': 'center', 'font_size': 11, 'valign': 'vcenter', 'bg_color': '#E4E5DF'})
                    style2 = workbook.add_format({'text_wrap': True, 'font_name': 'Calibri', 'border': 1, 'align': 'center', 'font_size': 11, 'valign': 'vcenter'})
                    ws.merge_range(0, 0, 0, 8, 'UNIVERSIDAD ESTATAL ESTATAL DE MILAGRO', title)
                    ws.merge_range(1, 1, 0, 8, 'LISTADO DE POSTULANTES MEJORES PUNTUADOS', title2)
                    row_num = 2
                    columns = [
                        ('Convocatoria', 90),
                        ('Categoria', 50),
                        ('Partida', 80),
                        ('Dedicacion', 80),
                        ('Cod. Unico', 15),
                        ('Apellidos', 30),
                        ('Nombres', 30),
                        ('Titulo', 50),
                        ('Archivo', 50),
                        ('Identificación', 20),
                        ('Correo', 30),
                        ('Telf.', 20),
                        ('Telf. Conv.', 20),
                        ('Tema', 100),
                        ('Fecha', 15),
                        ('Hora', 15),
                        ('Lugar', 25),
                        ('Obs', 60),
                        ('Nota Disertación.', 15),
                        ('Nota Entrevista', 15),
                        ('Nota Final', 15),
                        ('Revisado Por', 15),
                        ('Fecha Revisión', 40),
                    ]
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], fuentecabecera)
                        ws.set_column(col_num, col_num, columns[col_num][1])
                    row_num += 1
                    for partidas in Partida.objects.filter(status=True, convocatoria__vigente=True).order_by('convocatoria__descripcion'):
                        listado = partidas.personaaplicarpartida_set.filter(status=True,finsegundaetapa=True, nota_final__gte=69).order_by('-nota_final')[1:]
                        for det in listado:
                            if det.persona.mis_titulaciones().count()>0:
                                row_count = row_num + det.persona.mis_titulaciones().count()-1
                                ws.merge_range(row_num, 0,row_count,0 ,det.partida.convocatoria.descripcion, style2)
                                ws.merge_range(row_num, 1,row_count,1, det.partida.convocatoria.tipocontrato.nombre if det.partida.convocatoria.tipocontrato else '', style2)
                                ws.merge_range(row_num, 2,row_count,2, str(det.partida), style2)
                                ws.merge_range(row_num, 3,row_count,3, det.partida.get_dedicacion_display(), style2)
                                ws.merge_range(row_num, 4,row_count,4, det.pk, style2)
                                ws.merge_range(row_num, 5,row_count,5, "{} {}".format(det.persona.apellido1, det.persona.apellido2), style2)
                                ws.merge_range(row_num, 6,row_count,6, "{}".format(det.persona.nombres), style2)
                                ws.merge_range(row_num, 9,row_count,9, det.persona.cedula, style2)
                                ws.merge_range(row_num, 10,row_count,10, det.persona.email, style2)
                                ws.merge_range(row_num, 11,row_count,11, det.persona.telefono, style2)
                                ws.merge_range(row_num, 12,row_count,12, det.persona.telefono_conv, style2)
                                tema, fechaasistencia, horaasistencia, lugar, observacion = '', '', '', '', ''
                                if det.traer_agenda_entrevista():
                                    tema = det.traer_agenda_entrevista().tema
                                    fechaasistencia = det.traer_agenda_entrevista().fechaasistencia
                                    horaasistencia = det.traer_agenda_entrevista().horasistencia
                                    lugar = det.traer_agenda_entrevista().lugar
                                    observacion = det.traer_agenda_entrevista().observacion
                                ws.merge_range(row_num, 13,row_count,13, tema, style2)
                                ws.merge_range(row_num, 14,row_count,14, str(fechaasistencia), style2)
                                ws.merge_range(row_num, 15,row_count,15, str(horaasistencia), style2)
                                ws.merge_range(row_num, 16,row_count,16, lugar, style2)
                                ws.merge_range(row_num, 17,row_count,17, observacion, style2)
                                nota70, nota30 = '', ''
                                if det.traer_calificacion_disertacion():
                                    if det.traer_calificacion_disertacion().finalizada:
                                        nota70 = det.traer_calificacion_disertacion().nota_porcentual_70()
                                if det.traer_calificacion_entrevista():
                                    if det.traer_calificacion_entrevista().finalizada:
                                        nota30 = det.traer_calificacion_entrevista().nota_porcentual_30()
                                ws.merge_range(row_num, 18,row_count,18, nota70, style2)
                                ws.merge_range(row_num, 19,row_count,19, nota30, style2)
                                if det.finsegundaetapa:
                                    ws.merge_range(row_num, 20,row_count,20, det.total_segunda_etapa(), style2)
                                else:
                                    ws.merge_range(row_num, 20,row_count,20, '', style2)
                                revisado_por, fecha_revision = '', ''
                                if det.finsegundaetapa:
                                    revisado_por = det.setapa_revisado_por.username if det.setapa_revisado_por else ''
                                    fecha_revision = str(det.setapa_fecha_revision)
                                ws.merge_range(row_num, 21,row_count,21, revisado_por, style2)
                                ws.merge_range(row_num, 22,row_count,22, fecha_revision, style2)
                                for titulo in det.persona.mis_titulaciones():
                                    ws.write(row_num, 7, "{}".format(titulo.titulo), style2)
                                    ws.write(row_num, 8, "https://sga.unemi.edu.ec/{}".format(titulo.archivo.url) if titulo.archivo else '', style2)
                                    row_num +=1
                            else:
                                ws.write(row_num, 0, det.partida.convocatoria.descripcion, style2)
                                ws.write(row_num, 1, det.partida.convocatoria.tipocontrato.nombre if det.partida.convocatoria.tipocontrato else '', style2)
                                ws.write(row_num, 2, str(det.partida), style2)
                                ws.write(row_num, 3, det.partida.get_dedicacion_display(), style2)
                                ws.write(row_num, 4, det.pk, style2)
                                ws.write(row_num, 5, "{} {}".format(det.persona.apellido1, det.persona.apellido2), style2)
                                ws.write(row_num, 6, "{}".format(det.persona.nombres), style2)
                                ws.write(row_num, 7, "", style2)
                                ws.write(row_num, 8, "", style2)
                                ws.write(row_num, 9, det.persona.cedula, style2)
                                ws.write(row_num, 10, det.persona.email, style2)
                                ws.write(row_num,11, det.persona.telefono, style2)
                                ws.write(row_num, 12, det.persona.telefono_conv, style2)
                                tema, fechaasistencia, horaasistencia, lugar, observacion = '', '', '', '', ''
                                if det.traer_agenda_entrevista():
                                    tema = det.traer_agenda_entrevista().tema
                                    fechaasistencia = det.traer_agenda_entrevista().fechaasistencia
                                    horaasistencia = det.traer_agenda_entrevista().horasistencia
                                    lugar = det.traer_agenda_entrevista().lugar
                                    observacion = det.traer_agenda_entrevista().observacion
                                ws.write(row_num, 13, tema, style2)
                                ws.write(row_num, 14, str(fechaasistencia), style2)
                                ws.write(row_num, 15, str(horaasistencia), style2)
                                ws.write(row_num, 16, lugar, style2)
                                ws.write(row_num, 17, observacion, style2)
                                nota70, nota30 = '', ''
                                if det.traer_calificacion_disertacion():
                                    if det.traer_calificacion_disertacion().finalizada:
                                        nota70 = det.traer_calificacion_disertacion().nota_porcentual_70()
                                if det.traer_calificacion_entrevista():
                                    if det.traer_calificacion_entrevista().finalizada:
                                        nota30 = det.traer_calificacion_entrevista().nota_porcentual_30()
                                ws.write(row_num, 18, nota70, style2)
                                ws.write(row_num, 19, nota30, style2)
                                if det.finsegundaetapa:
                                    ws.write(row_num, 20, det.total_segunda_etapa(), style2)
                                else:
                                    ws.write(row_num, 20, '', style2)
                                revisado_por, fecha_revision = '', ''
                                if det.finsegundaetapa:
                                    revisado_por = det.setapa_revisado_por.username if det.setapa_revisado_por else ''
                                    fecha_revision = str(det.setapa_fecha_revision)
                                ws.write(row_num, 21, revisado_por, style2)
                                ws.write(row_num, 22, fecha_revision, style2)
                                row_num += 1
                    workbook.close()
                    output.seek(0)
                    filename = 'reporte_sin_mejores_puntuados_' + random.randint(1, 10000).__str__() + '.xlsx'
                    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    response['Content-Disposition'] = 'attachment; filename=%s' % filename
                    return response
                except Exception as ex:
                    return "{}?info=ErrorLinea{}Error".format(request.path, str(sys.exc_info()[-1].tb_lineno), str(ex))

            if action == 'establecerganador':
                try:
                    data['id'] = id = int(request.GET['id'])
                    data['partida'] = partida = Partida.objects.get(status = True,id=int(request.GET['id']))
                    nummejores = partida.convocatoria.nummejorespuntuados
                    data['postulantes'] = partida.personaaplicarpartida_set.filter(status = True,estado__in=[1,4],finsegundaetapa=True).order_by('-nota_final')[:nummejores]
                    if partida.convocatoria.modeloevaluativoconvocatoria:
                        data['postulantes'] = partida.personaaplicarpartida_set.filter(status = True,estado__in=[4],finsegundaetapa=True).order_by('-nota_final')[:nummejores]
                    template = get_template('postulate/adm_segundaetapa/estaganador.html')
                    return JsonResponse({'result':True,'data':template.render(data)})
                except Exception as ex:
                    pass

            if action == 'consultaarchivo':
                try:
                    persona = PersonaAplicarPartida.objects.get(pk=request.GET['id'])
                    modelo = persona.calificaciondisertacion_set.filter(status=True)[0]
                    calificaciondisertacion = CalificacionDisertacion.objects.filter(status=True, postulacion=persona, modeloevaluativo=modelo.modeloevaluativo)[0]
                    # data['partida'] = partida = Partida.objects.get(status = True,id=int(request.GET['id']))
                    # data['postulantes'] = partida.personaaplicarpartida_set.filter(status = True,estado = 1,finsegundaetapa=True).order_by('-nota_final')[:3]
                    # template = get_template('postulate/adm_segundaetapa/estaganador.html')
                    return JsonResponse({'result':True,'modelo':calificaciondisertacion.archivo.url if calificaciondisertacion.archivo else ''},safe=False)
                except Exception as ex:
                    pass

            elif action == 'importarnotas':
                try:
                    id = int(request.GET['id'])
                    data['id'] = id
                    convocatoria = Convocatoria.objects.get(status=True,id=id)
                    if not convocatoria.modeloevaluativoconvocatoria:
                        raise NameError("La convocatoria seleccionada no tiene configurado un modelo evaluativo convocatoria")
                    data['title'] = u'Importar notas'
                    data['mensaje'] = True
                    data['rutaarchivo'] = '/static/formatos/formato_carga_notas.xlsx'
                    form = SubirArchivoForm()
                    form.fields['campo'].queryset = DetalleModeloEvaluativoConvocatoria.objects.filter(status=True,modelo=convocatoria.modeloevaluativoconvocatoria)
                    data['form'] = form
                    template = get_template("postulate/adm_segundaetapa/modal/archivo.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    err_ = f"Ocurrio un error: {ex.__str__()}. En la linea {sys.exc_info()[-1].tb_lineno}"
                    return JsonResponse({'result': False, 'mensaje': err_})

            elif action == 'loadfields':
                try:
                    id = request.GET['id']
                    registro = ModeloEvaluativoConvocatoria.objects.get(status=True,id=int(id))
                    detalle = DetalleModeloEvaluativoConvocatoria.objects.filter(status=True,modelo=registro).order_by('-nombre')
                    lista = []
                    for dt in detalle:
                        lista.append([dt.id, dt.__str__()])
                    res_js = {'result':True,'lista':lista}
                except Exception as ex:
                    err_ = f"Ocurrio un error: {ex.__str__()}. En la linea {sys.exc_info()[-1].tb_lineno}"
                    res_js = {'result':False,'mensaje':err_}
                return JsonResponse(res_js)

            elif action == 'loadtestpsc':
                try:
                    data['title'] = u'Importar notas'
                    id = encrypt(request.GET['id'])
                    data['filtro'] = registro = PersonaAplicarPartida.objects.get(status=True,id=int(id))
                    form = SubirArchivoTestPsicologicoForm(initial=model_to_dict(registro))
                    data['form'] = form
                    template = get_template("postulate/adm_segundaetapa/modal/archivo.html")
                    res_js = {'result':True,'data':template.render(data)}
                except Exception as ex:
                    err_ = f"Ocurrio un error: {ex.__str__()}. En la linea {sys.exc_info()[-1].tb_lineno}"
                    res_js = {'result':False,'mensaje':err_}
                return JsonResponse(res_js)

            elif action == 'loadnotificacionentrevista':
                try:
                    data['title'] = u'Importar notas'
                    id = encrypt(request.GET['id'])
                    data['filtro'] = registro = PersonaAplicarPartida.objects.get(status=True, id=int(id))
                    template = get_template("postulate/adm_segundaetapa/modal/emailnotificacion.html")
                    res_js = {'result': True, 'data': template.render(data)}
                except Exception as ex:
                    err_ = f"Ocurrio un error: {ex.__str__()}({sys.exc_info()[-1].tb_lineno})"
                    res_js = {'result': False, 'mensaje': err_}
                return JsonResponse(res_js)

            elif action == 'loadnotificacionentr3pt':
                try:
                    data['title'] = u'Importar notas'
                    data['titulo_mail'] = "Fase Entrevista"
                    id = encrypt(request.GET['id'])
                    data['filtro'] = registro = Partida.objects.get(status=True, id=int(id))
                    template = get_template("postulate/adm_segundaetapa/modal/emailnotificacion.html")
                    res_js = {'result': True, 'data': template.render(data)}
                except Exception as ex:
                    err_ = f"Ocurrio un error: {ex.__str__()}({sys.exc_info()[-1].tb_lineno})"
                    res_js = {'result': False, 'mensaje': err_}
                return JsonResponse(res_js)

            elif action == 'loadnotificaciomenor70pt':
                try:
                    data['title'] = u'Importar notas'
                    data['titulo_mail'] = "Resultado de la Prueba Técnica"
                    id = encrypt(request.GET['id'])
                    data['filtro'] = registro = Partida.objects.get(status=True, id=int(id))
                    template = get_template("postulate/adm_segundaetapa/modal/emailnotificacion.html")
                    res_js = {'result': True, 'data': template.render(data)}
                except Exception as ex:
                    err_ = f"Ocurrio un error: {ex.__str__()}({sys.exc_info()[-1].tb_lineno})"
                    res_js = {'result': False, 'mensaje': err_}
                return JsonResponse(res_js)

            elif action == 'loadnotificacionentrexcl3pt':
                try:
                    data['title'] = u'Importar notas'
                    data['titulo_mail'] = "Resultado de la Prueba Técnica - No Seleccionado"
                    id = encrypt(request.GET['id'])
                    data['filtro'] = registro = Partida.objects.get(status=True, id=int(id))
                    template = get_template("postulate/adm_segundaetapa/modal/emailnotificacion.html")
                    res_js = {'result': True, 'data': template.render(data)}
                except Exception as ex:
                    err_ = f"Ocurrio un error: {ex.__str__()}({sys.exc_info()[-1].tb_lineno})"
                    res_js = {'result': False, 'mensaje': err_}
                return JsonResponse(res_js)

            elif action == 'loadnotificacionotafinalupper':
                try:
                    data['title'] = u'Importar notas'
                    data['titulo_mail'] = "Resultado de las Entrevistas Finales"
                    id = encrypt(request.GET['id'])
                    data['filtro'] = registro = Partida.objects.get(status=True, id=int(id))
                    template = get_template("postulate/adm_segundaetapa/modal/emailnotificacion.html")
                    res_js = {'result': True, 'data': template.render(data)}
                except Exception as ex:
                    err_ = f"Ocurrio un error: {ex.__str__()}({sys.exc_info()[-1].tb_lineno})"
                    res_js = {'result': False, 'mensaje': err_}
                return JsonResponse(res_js)

            elif action == 'loadnotificacionotafinallower':
                try:
                    data['title'] = u'Importar notas'
                    data['titulo_mail'] = "Resultado del Proceso de Selección"
                    id = encrypt(request.GET['id'])
                    data['filtro'] = registro = Partida.objects.get(status=True, id=int(id))
                    template = get_template("postulate/adm_segundaetapa/modal/emailnotificacion.html")
                    res_js = {'result': True, 'data': template.render(data)}
                except Exception as ex:
                    err_ = f"Ocurrio un error: {ex.__str__()}({sys.exc_info()[-1].tb_lineno})"
                    res_js = {'result': False, 'mensaje': err_}
                return JsonResponse(res_js)

            elif action == 'view3pt':
                try:
                    data['title'] = 'Mejores puntuados (PT)'
                    data['actionmodal'] = 'loadnotificacionentr3pt'
                    id = encrypt(request.GET.get('id'))
                    data['partida'] = partida = Partida.objects.get(status=True, id=int(id))
                    persoapartida = PersonaAplicarPartida.objects.filter(status=True,partida=partida,evaluacionpostulante__isnull=False,
                                                                 evaluacionpostulante__detallemodeloevaluativo__nombre='PT',
                                                                 evaluacionpostulante__valor__gte=70).order_by('-evaluacionpostulante__valor')[:3]
                    data['lista'] = persoapartida
                    return render(request,'postulate/adm_segundaetapa/view3pt.html',data)
                except Exception as ex:
                    err_ = f"Ocurrio un error: {ex.__str__()}({sys.exc_info()[-1].tb_lineno})"
                    return HttpResponseRedirect(f'{request.path}?info={err_}')

            elif action == 'viewptexcl3pt':
                try:
                    data['title'] = 'Mayor a 70 excluyendo los mejores 3 (Prueba Técnica)'
                    data['actionmodal'] = 'loadnotificacionentrexcl3pt'
                    id = encrypt(request.GET.get('id'))
                    data['partida'] = partida = Partida.objects.get(status=True, id=int(id))
                    persoapartida = PersonaAplicarPartida.objects.filter(status=True,partida=partida,evaluacionpostulante__isnull=False,
                                                                 evaluacionpostulante__detallemodeloevaluativo__nombre='PT',
                                                                 evaluacionpostulante__valor__gte=70).order_by('-evaluacionpostulante__valor')[3:]
                    data['lista'] = persoapartida
                    return render(request,'postulate/adm_segundaetapa/view3pt.html',data)
                except Exception as ex:
                    err_ = f"Ocurrio un error: {ex.__str__()}({sys.exc_info()[-1].tb_lineno})"
                    return HttpResponseRedirect(f'{request.path}?info={err_}')

            elif action == 'viewmenor70pt':
                try:
                    data['title'] = 'Menor a 70 (Prueba Técnica)'
                    id = encrypt(request.GET.get('id'))
                    data['actionmodal'] = 'loadnotificaciomenor70pt'
                    data['partida'] = partida = Partida.objects.get(status=True, id=int(id))
                    persoapartida = PersonaAplicarPartida.objects.filter(status=True,partida=partida,evaluacionpostulante__isnull=False,
                                                                 evaluacionpostulante__detallemodeloevaluativo__nombre='PT',
                                                                 evaluacionpostulante__valor__lt=70).order_by('-evaluacionpostulante__valor')
                    data['lista'] = persoapartida
                    return render(request,'postulate/adm_segundaetapa/view3pt.html',data)
                except Exception as ex:
                    err_ = f"Ocurrio un error: {ex.__str__()}({sys.exc_info()[-1].tb_lineno})"
                    return HttpResponseRedirect(f'{request.path}?info={err_}')

            elif action == 'viewnotafinalupper':
                try:
                    data['title'] = 'Mayor igual a 70 nota final excluyendo al ganador (Nota Final)'
                    id = encrypt(request.GET.get('id'))
                    data['actionmodal'] = 'loadnotificacionotafinalupper'
                    data['partida'] = partida = Partida.objects.get(status=True, id=int(id))
                    persoapartida = PersonaAplicarPartida.objects.filter(status=True,partida=partida,nota_final__gte=70).exclude(esganador=True)
                    data['lista'] = persoapartida
                    return render(request,'postulate/adm_segundaetapa/view3pt.html',data)
                except Exception as ex:
                    err_ = f"Ocurrio un error: {ex.__str__()}({sys.exc_info()[-1].tb_lineno})"
                    return HttpResponseRedirect(f'{request.path}?info={err_}')

            elif action == 'viewnotafinallower':
                try:
                    data['title'] = 'Menor a 70 nota final (Nota Final)'
                    id = encrypt(request.GET.get('id'))
                    data['actionmodal'] = 'loadnotificacionotafinallower'
                    data['partida'] = partida = Partida.objects.get(status=True, id=int(id))
                    persoapartida = PersonaAplicarPartida.objects.filter(status=True,partida=partida,nota_final__lt=70,finsegundaetapa=True).exclude(esganador=True)
                    data['lista'] = persoapartida
                    return render(request,'postulate/adm_segundaetapa/view3pt.html',data)
                except Exception as ex:
                    err_ = f"Ocurrio un error: {ex.__str__()}({sys.exc_info()[-1].tb_lineno})"
                    return HttpResponseRedirect(f'{request.path}?info={err_}')

        else:
            try:
                data['title'] = u'Disertación/Entrevista Postulaciones'
                qsconvocatorias = Convocatoria.objects.filter(status=True, segundaetapa=True)
                validarpartidas = True
                departamentogestion = Departamento.objects.filter(status=True, permisodepartamento=2)
                if departamentogestion.exists():
                    if departamentogestion.first().mis_integrantes().filter(id=persona.pk).exists():
                        validarpartidas = False
                if request.user.is_superuser:
                    validarpartidas = False
                if validarpartidas:
                    idsconvocatorias = PartidaTribunal.objects.filter(status=True, persona=persona, tipo=2).values_list('partida__convocatoria__id', flat=True)
                    qsconvocatorias = qsconvocatorias.filter(id__in=idsconvocatorias)
                data['convocatorias'] = qsconvocatorias.order_by('-id')
                return render(request, "postulate/adm_segundaetapa/view.html", data)
            except Exception as ex:
                pass

def actualizar_nota_postulante_2(request, postulante, sel, valor):
    postulante = PersonaAplicarPartida.objects.get(pk=postulante.id)
    modeloevaluativo = postulante.partida.convocatoria.modeloevaluativoconvocatoria
    campomodelo = modeloevaluativo.campo(sel)
    try:
        if not valor:
            valor = null_to_decimal(float(valor), campomodelo.decimales)
        if valor >= campomodelo.notamaxima:
            valor = campomodelo.notamaxima
        elif valor <= campomodelo.notaminima:
            valor = campomodelo.notaminima
    except:
        valor = campomodelo.notaminima
    campo = postulante.campo(sel)
    campo.valor = valor
    campo.save(request)
    d = locals()
    exec(modeloevaluativo.logicamodelo, globals(), d)
    d['calculo_modelo_evaluativo'](postulante)
    postulante.nota_final = null_to_decimal(postulante.nota_final, 2)
    if postulante.nota_final > modeloevaluativo.notamaxima:
        postulante.nota_final = modeloevaluativo.notamaxima
    postulante.save(request)
    camposdependientes = []
    for campomodelo in modeloevaluativo.campos():
        if campomodelo.dependiente:
            camposdependientes.append((campomodelo.htmlid(), postulante.valor_nombre_campo(campomodelo.nombre), campomodelo.decimales))

    postulante.actualiza_estado()
    postulante.actualiza_estadofinal()
    postulante.save(request)

def envio_email_pt_nopasaron(**kwargs):
    """
    Función para enviar al correo a las personas que han rendido la prueba técnica que no alcanzaro el puntaje mínimo
    de 70 puntos
    :param kwargs:
    :return:
    """
    try:
        request = kwargs.get('request', None)
        _var = lambda x: request.session.get(x, None) if not x in kwargs else kwargs.get(x, None)

        convocatoria = _var('msg_convocatoria')
        titulo = _var('msg_title')
        contenido = _var('msg_contenido')
        contenido_static = _var('msg_contenido_header')
        conv = Convocatoria.objects.get(status=True, id=int(convocatoria))
        partidas = conv.partida_set.filter(status=True)
        for part in partidas:
            peraplicapart = PersonaAplicarPartida.objects.filter(status=True, partida=part,nota_final__gte=70, esganador=True)
            cont = 0
            for perap in peraplicapart:
                cont += 1
                personaemail_ = []
                if perap.persona.email:
                    personaemail_.append(perap.persona.email)
                contenido_static = f"""<div  class="mb-1" style="text-align: center;">
                        <div id="mensaje_head"><br>Respecto a su aplicación en el Proceso de Preselección para la partida <b id="nombre_convocatoria">{part.convocatoria}</b> convocado por la INSTITUCIÓN:<b>UNIVERSIDAD ESTATAL DE MILAGRO</b>
                        </div>
                        </div>
                """
                if personaemail_:
                    send_html_mail(f"Postulate: {titulo}",
                                   "emails/segundaetapa_postulate.html",
                                   {'sistema': u'SISTEMA POSTULATE UNEMI', 'persona': perap.persona,
                                    'partida': part, 'contenido': contenido, 'contenido_static':contenido_static,
                                    'title':titulo,
                                    't': miinstitucion()}, [personaemail_], [],
                                   cuenta=CUENTAS_CORREOS[30][1])
                if not NotificacionGanador.objects.filter(status=True,personaaplicapartida=perap):
                    notificar = NotificacionGanador(personaaplicapartida=perap,
                                                    fecha=datetime.now().date())
                    notificar.save()
            logobjeto(u'Notificación Finalización de la prueba técnica: %s' % part, request, "add", None, part)
    except Exception as ex:
        pass

def envio_email_pt_pasaron(**kwargs):
    """
    Función para enviar al correo a las personas que han rendido la prueba técnica y obtuvieron la nota mínima de 70 pero
    no quedaron entre los 3 primeros con mayor puntaje.
    :param kwargs:
    :return:
    """
    try:
        request = kwargs.get('request', None)
        _var = lambda x: request.session.get(x, None) if not x in kwargs else kwargs.get(x, None)

        convocatoria = _var('msg_convocatoria')
        titulo = _var('msg_title')
        contenido = _var('msg_contenido')
        contenido_static = _var('msg_contenido_header')
        conv = Convocatoria.objects.get(status=True, id=int(convocatoria))
        partidas = conv.partida_set.filter(status=True)
        for part in partidas:
            top_3_scores = (
                PersonaAplicarPartida.objects
                .filter(
                    status=True,
                    partida=part,
                    evaluacionpostulante__isnull=False,
                    evaluacionpostulante__detallemodeloevaluativo__nombre='PT',
                    evaluacionpostulante__valor__gte=70
                )
                .order_by('-evaluacionpostulante__valor')
                .values_list('evaluacionpostulante__valor', flat=True)[:3]
            )
            peraplicapart = PersonaAplicarPartida.objects.filter(status=True, partida=part,
                                                                 evaluacionpostulante__isnull=False,
                                                                 evaluacionpostulante__detallemodeloevaluativo__nombre='PT',
                                                                 evaluacionpostulante__valor__gte=70).exclude(
                evaluacionpostulante__valor__in=top_3_scores
            )
            cont = 0
            for perap in peraplicapart:
                cont += 1
                personaemail_ = []
                if perap.persona.email:
                    personaemail_.append(perap.persona.email)
                if personaemail_:
                    send_html_mail(f"Postulate: {titulo}",
                                   "emails/segundaetapa_postulate.html",
                                   {'sistema': u'SISTEMA POSTULATE UNEMI', 'persona': perap.persona,
                                    'partida': part, 'contenido': contenido, 'contenido_static': contenido_static,
                                    'title': titulo,
                                    't': miinstitucion()}, [personaemail_], [],
                                   cuenta=CUENTAS_CORREOS[30][1])
            logobjeto(u'Notificación Finalización de la prueba técnica: %s' % part, request, "add", None, part)
    except Exception as ex:
        pass

def envio_email_notafinal_exclu_ganador(**kwargs):
    """
       Función para enviar al correo a las personas cuando ya se ha establecido el ganador de las partidas
       :param kwargs:
       :return:
   """
    try:
        pass
    except Exception as ex:
        pass

def envio_email_3mejores_pt(**kwargs):
    try:
        request = kwargs.get('request', None)
        _var = lambda x: request.session.get(x, None) if not x in kwargs else kwargs.get(x, None)

        partida = _var('partida')
        titulo = _var('msg_title')
        contenido = _var('msg_contenido')
        contenido_static = _var('msg_contenido_header')
        partida = partida
        postulantes = PersonaAplicarPartida.objects.filter(status=True,partida=partida, evaluacionpostulante__isnull=False,
                                                                 evaluacionpostulante__detallemodeloevaluativo__nombre='PT',
                                                                 evaluacionpostulante__valor__gte=70).order_by('-evaluacionpostulante__valor')[:3]
        cont = 0
        for pst in postulantes:
            personaemail_ = []
            if pst.persona.email:
                personaemail_.append(pst.persona.email)
            if cont<1:
                personaemail_.append('rviterib1@unemi.edu.ec')
            cont +=1
            if personaemail_:
                send_html_mail(f"Postulate: {titulo}",
                               "emails/segundaetapa_postulate.html",
                               {'sistema': u'SISTEMA POSTULATE UNEMI', 'persona': pst.persona,
                                'partida': partida, 'contenido': contenido, 'contenido_static': contenido_static,
                                'title': titulo,
                                't': miinstitucion()}, [personaemail_], [],
                               cuenta=CUENTAS_CORREOS[30][1])
                time.sleep(3)

    except Exception as ex:
        pass

def envio_email_excl3mejores_pt(**kwargs):
    try:
        request = kwargs.get('request', None)
        _var = lambda x: request.session.get(x, None) if not x in kwargs else kwargs.get(x, None)

        partida = _var('partida')
        titulo = _var('msg_title')
        contenido = _var('msg_contenido')
        contenido_static = _var('msg_contenido_header')
        partida = partida
        postulantes = PersonaAplicarPartida.objects.filter(status=True,partida=partida, evaluacionpostulante__isnull=False,
                                                                 evaluacionpostulante__detallemodeloevaluativo__nombre='PT',
                                                                 evaluacionpostulante__valor__gte=70).order_by('-evaluacionpostulante__valor')[3:]
        cont = 0
        for pst in postulantes:
            personaemail_ = []
            if pst.persona.email:
                personaemail_.append(pst.persona.email)
            if cont<1:
                personaemail_.append('rviterib1@unemi.edu.ec')
            cont +=1
            if personaemail_:
                send_html_mail(f"Postulate: {titulo}",
                               "emails/segundaetapa_postulate.html",
                               {'sistema': u'SISTEMA POSTULATE UNEMI', 'persona': pst.persona,
                                'partida': partida, 'contenido': contenido, 'contenido_static': contenido_static,
                                'title': titulo,
                                't': miinstitucion()}, [personaemail_], [],
                               cuenta=CUENTAS_CORREOS[30][1])
                time.sleep(3)
    except Exception as ex:
        pass

def envio_email_m70_pt(**kwargs):
    try:
        request = kwargs.get('request', None)
        _var = lambda x: request.session.get(x, None) if not x in kwargs else kwargs.get(x, None)

        partida = _var('partida')
        titulo = _var('msg_title')
        contenido = _var('msg_contenido')
        contenido_static = _var('msg_contenido_header')
        partida = partida
        postulantes = PersonaAplicarPartida.objects.filter(status=True,partida=partida, evaluacionpostulante__isnull=False,
                                                                 evaluacionpostulante__detallemodeloevaluativo__nombre='PT',
                                                                 evaluacionpostulante__valor__lt=70).order_by('-evaluacionpostulante__valor')
        cont = 0
        for pst in postulantes:
            personaemail_ = []
            if cont<1:
                personaemail_.append('rviterib1@unemi.edu.ec')
            cont +=1
            if pst.persona.email:
                personaemail_.append(pst.persona.email)
            if personaemail_:
                send_html_mail(f"Postulate: {titulo}",
                               "emails/segundaetapa_postulate.html",
                               {'sistema': u'SISTEMA POSTULATE UNEMI', 'persona': pst.persona,
                                'partida': partida, 'contenido': contenido, 'contenido_static': contenido_static,
                                'title': titulo,
                                't': miinstitucion()}, [personaemail_], [],
                               cuenta=CUENTAS_CORREOS[30][1])
                time.sleep(3)
    except Exception as ex:
        pass

def envio_email_notafina_70up(**kwargs):
    try:
        request = kwargs.get('request', None)
        _var = lambda x: request.session.get(x, None) if not x in kwargs else kwargs.get(x, None)

        partida = _var('partida')
        titulo = _var('msg_title')
        contenido = _var('msg_contenido')
        contenido_static = _var('msg_contenido_header')
        partida = partida
        postulantes = PersonaAplicarPartida.objects.filter(status=True,partida=partida, nota_final__gte=70).exclude(esganador=True).order_by('-nota_final')
        cont = 0
        for pst in postulantes:
            personaemail_ = []
            if cont<1:
                personaemail_.append('rviterib1@unemi.edu.ec')
            cont +=1
            if pst.persona.email:
                personaemail_.append(pst.persona.email)
            if personaemail_:
                send_html_mail(f"Postulate: {titulo}",
                               "emails/segundaetapa_postulate.html",
                               {'sistema': u'SISTEMA POSTULATE UNEMI', 'persona': pst.persona,
                                'partida': partida, 'contenido': contenido, 'contenido_static': contenido_static,
                                'title': titulo,
                                't': miinstitucion()}, [personaemail_], [],
                               cuenta=CUENTAS_CORREOS[30][1])
                time.sleep(3)
    except Exception as ex:
        pass

def envio_email_notafina_70down(**kwargs):
    try:
        request = kwargs.get('request', None)
        _var = lambda x: request.session.get(x, None) if not x in kwargs else kwargs.get(x, None)

        partida = _var('partida')
        titulo = _var('msg_title')
        contenido = _var('msg_contenido')
        contenido_static = _var('msg_contenido_header')
        partida = partida
        postulantes = PersonaAplicarPartida.objects.filter(status=True,partida=partida,finsegundaetapa=True, nota_final__lt=70).order_by('-nota_final')
        cont = 0
        for pst in postulantes:
            personaemail_ = []
            if cont<1:
                personaemail_.append('rviterib1@unemi.edu.ec')
            cont +=1
            if pst.persona.email:
                personaemail_.append(pst.persona.email)
            if personaemail_:
                send_html_mail(f"Postulate: {titulo}",
                               "emails/segundaetapa_postulate.html",
                               {'sistema': u'SISTEMA POSTULATE UNEMI', 'persona': pst.persona,
                                'partida': partida, 'contenido': contenido, 'contenido_static': contenido_static,
                                'title': titulo,
                                't': miinstitucion()}, [personaemail_], [],
                               cuenta=CUENTAS_CORREOS[30][1])
                time.sleep(3)
    except Exception as ex:
        pass