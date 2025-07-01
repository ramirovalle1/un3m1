#!/usr/bin/env python
# -*- coding: utf-8 -*-
import os
import sys
import openpyxl
import io
import subprocess
import xlsxwriter
from django.http import HttpResponseRedirect, JsonResponse, HttpResponse
# SITE_ROOT = os.path.dirname(os.path.realpath(__file__))
YOUR_PATH = os.path.dirname(os.path.realpath(__file__))
# print(f"YOUR_PATH: {YOUR_PATH}")
SITE_ROOT = os.path.dirname(os.path.dirname(YOUR_PATH))
SITE_ROOT = os.path.join(SITE_ROOT, '')
# print(f"SITE_ROOT: {SITE_ROOT}")
your_djangoproject_home = os.path.split(SITE_ROOT)[0]
# print(f"your_djangoproject_home: {your_djangoproject_home}")
sys.path.append(your_djangoproject_home)
import pyqrcode
from django.core.wsgi import get_wsgi_application
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "settings")
application = get_wsgi_application()
from sga.models import *
from sagest.models import *
from django.core.files.base import ContentFile
from django.template import engines
from xhtml2pdf import pisa
import io as StringIO
from sga.funciones import log
from sga.reportes import run_report_v1
# from balcon.models import *
from moodle import moodle
from posgrado.models import InscripcionCohorte, AsesorComercial, HistorialAsesor, \
    EvidenciaRequisitosAspirante, RequisitosMaestria, TipoRespuestaProspecto, \
    DetalleAprobacionFormaPago, DetalleEvidenciaRequisitosAspirante, VentasProgramaMaestria, \
    AsesorMeta, MaestriasAdmision, ConfigFinanciamientoCohorte
from sga.commonviews import ficha_socioeconomica
from Moodle_Funciones import crearhtmlphpmoodle
from inno.models import TipoFormaPagoPac
import warnings
from certi.models import Certificado
from sga.templatetags.sga_extras import encrypt
from datetime import datetime, time
from settings import SITE_STORAGE, MEDIA_ROOT, MEDIA_URL
from django.contrib.contenttypes.fields import ContentType
from secretaria.models import *
from secretaria.funciones import generar_codigo_solicitud
import shutil
from api.helpers.functions_helper import get_variable, remove_accents, fixparametro, fetch_resources
from settings import EMAIL_DOMAIN, DEBUG, JR_USEROUTPUT_FOLDER, MEDIA_URL, SITE_STORAGE, JR_JAVA_COMMAND, JR_RUN, \
    DATABASES, MEDIA_ROOT, SUBREPOTRS_FOLDER
from sga.funciones import log, generar_codigo, variable_valor, generar_nombre
from certi.models import Certificado, CertificadoAsistenteCertificadora, CertificadoUnidadCertificadora
from secretaria.funciones import ReportBackground, transform_jasperstarter
from settings import USA_TIPOS_INSCRIPCIONES, TIPO_INSCRIPCION_INICIAL
import zipfile

warnings.filterwarnings('ignore', message='Unverified HTTPS request')
print(u"Inicio")
print(u"********************Cerrar materia ***************************")

try:
    c = 0
    eMaterias = Materia.objects.filter(status=True, nivel__periodo__id=317,
                                       asignaturamalla__malla__carrera__coordinacion__id__in=[1, 2, 3, 4, 5])
    for eMateria in eMaterias:
        debe_cerrar_periodo = True
        if eMateria.nivel.periodo.periodo_academia():
            if not eMateria.nivel.periodo.periodo_academia().cierra_materia:
                debe_cerrar_periodo = False
                raise Exception('En este periodo no debe cerrarse las materias.')
        if debe_cerrar_periodo:
            eMateria.cerrado = True
            eMateria.fechacierre = datetime.now().date()
            eMateria.save()

            for asig in eMateria.asignados_a_esta_materia():
                asig.cerrado = True
                asig.save(actualiza=False)
                asig.actualiza_estado()
            for asig in eMateria.asignados_a_esta_materia():
                asig.cierre_materia_asignada()
        c += 1
        print(f'{c}/{eMaterias.count()} - {eMateria.id} - {eMateria.asignaturamalla.asignatura.nombre} - {eMateria.paralelo}')
except Exception as ex:
    pass
# def generar_certificado_digital(ePersona, eReporte, eUser, params, eSolicitud=None):
#     data = {}
#     cambiaruta = 0
#     isQR = False
#     codigo = None
#     certificado = None
#     base_url = 'https://sga.unemi.edu.ec/'
#     app = 'sga'
#     if 'app' in params:
#         app = params['app']
#     if eReporte.archivo:
#         tipo = params['rt']
#         output_folder = os.path.join(JR_USEROUTPUT_FOLDER, remove_accents(eUser.username))
#         try:
#             os.makedirs(output_folder)
#         except Exception as ex:
#             pass
#         d = datetime.now()
#         pdfname = eReporte.nombre + d.strftime('%Y%m%d_%H%M%S')
#         if eReporte.version == 2:
#             """SE REALIZA AJUSTES PARA NUEVAS VERSIONES"""
#             content_type = None
#             object_id = None
#             vqr = None
#             if 'vqr' in params:
#                 isQR = True
#                 vqr = params['vqr']
#                 # pdfname = reporte.nombre + vqr
#                 parametros_reporte = eReporte.parametroreporte_set.get(nombre='vqr')
#                 if not Certificado.objects.filter(reporte=eReporte).exists():
#                     raise NameError(u"No se encontro certificado activo" if eUser.is_superuser else u"Código: ADM_001")
#                 if Certificado.objects.filter(reporte=eReporte).count() > 1:
#                     raise NameError(u"Mas de un certificado activo" if eUser.is_superuser else u"Código: ADM_002")
#                 certificado = Certificado.objects.get(reporte=eReporte)
#
#                 if parametros_reporte.extra:
#                     try:
#                         content_type = ContentType.objects.get_for_model(eval(parametros_reporte.extra))
#                         object_id = vqr
#                     except Exception as ex:
#                         content_type = ContentType.objects.get_for_model(eUser)
#                         object_id = eUser.id
#                 else:
#                     try:
#                         if Persona.objects.filter(usuario_id=eUser.id).exists():
#                             persona = Persona.objects.filter(usuario_id=eUser.id).first()
#                             content_type = ContentType.objects.get_for_model(persona)
#                             object_id = persona.id
#                         else:
#                             content_type = ContentType.objects.get_for_model(eUser)
#                             object_id = eUser.id
#                     except Exception as ex:
#                         content_type = ContentType.objects.get_for_model(eUser)
#                         object_id = eUser.id
#             else:
#                 try:
#                     if Persona.objects.filter(usuario_id=eUser.id).exists():
#                         persona = Persona.objects.filter(usuario_id=eUser.id).first()
#                         content_type = ContentType.objects.get_for_model(persona)
#                         object_id = persona.id
#                     else:
#                         content_type = ContentType.objects.get_for_model(eUser)
#                         object_id = eUser.id
#                 except Exception as ex:
#                     content_type = ContentType.objects.get_for_model(eUser)
#                     object_id = eUser.id
#
#             if isQR:
#                 SUFFIX = None
#                 if (parametros_reporte.extra).upper() == "MATRICULA":
#                     # print(vqr, encrypt(vqr))
#                     matricula = Matricula.objects.get(pk=encrypt(vqr))
#                     carrera = matricula.inscripcion.carrera
#                     coordinacion = matricula.inscripcion.coordinacion
#                 else:
#                     inscripcion = Inscripcion.objects.get(pk=encrypt(vqr))
#                     carrera = inscripcion.carrera
#                     coordinacion = inscripcion.coordinacion
#                 if not certificado:
#                     raise NameError(u"No se encontro certificado." if eUser.is_superuser else u"Código: ADM_003")
#                 uc = None
#                 ac = None
#                 if not certificado.tiene_unidades_certificadoras():
#                     raise NameError(u"Certificado no tiene configurado Unidad certificadora" if eUser.is_superuser else u"Código: ADM_004")
#                 if certificado.tipo_origen == 1 and certificado.tipo_validacion == 2:
#                     if not CertificadoAsistenteCertificadora.objects.filter(status=True, carrera=carrera, unidad_certificadora__certificado=certificado, unidad_certificadora__coordinacion=coordinacion).exists():
#                         raise NameError(u"Certificado no tiene configurado Asistente certificadora" if eUser.is_superuser else u"Código: ADM_005")
#                     ac = CertificadoAsistenteCertificadora.objects.get(status=True, carrera=carrera, unidad_certificadora__certificado=certificado, unidad_certificadora__coordinacion=coordinacion)
#                     uc = ac.unidad_certificadora
#                 else:
#                     if not certificado.unidades_certificadoras().count() > 0:
#                         raise NameError(u"Certificado tiene configurado mas de una Unidad certificadora" if eUser.is_superuser else u"Código: ADM_006")
#                     uc = CertificadoUnidadCertificadora.objects.get(certificado=certificado)
#                 if not uc or not uc.alias:
#                     raise NameError(u"Certificado no tiene configurado Unidad certificadora" if eUser.is_superuser else u"Código: ADM_007")
#                 SUFFIX = uc.alias
#                 secuencia = 1
#                 try:
#                     if not LogReporteDescarga.objects.filter(fechahora__year=datetime.now().year, suffix=SUFFIX, secuencia__gt=0).exists():
#                         secuencia = 1
#                     else:
#                         if LogReporteDescarga.objects.filter(secuencia__gt=0, suffix=SUFFIX).order_by("-secuencia").exists():
#                             sec = LogReporteDescarga.objects.filter(secuencia__gt=0, suffix=SUFFIX).order_by("-secuencia")[0].secuencia
#                             if sec:
#                                 secuencia = int(sec) + 1
#                 except:
#                     pass
#                 codigo = generar_codigo(secuencia, 'UNEMI', SUFFIX, 7)
#
#             url = "/".join([MEDIA_URL, 'documentos', 'userreports', remove_accents(eUser.username), pdfname + "." + tipo])
#             logreporte = LogReporteDescarga(reporte=eReporte,
#                                             content_type=content_type,
#                                             object_id=encrypt(object_id),
#                                             url=url,
#                                             fechahora=datetime.now())
#             logreporte.save(usuario_id=eUser.id)
#             if isQR:
#                 logreporte.secuencia = secuencia
#                 logreporte.codigo = codigo
#                 logreporte.prefix = 'UNEMI'
#                 logreporte.suffix = SUFFIX
#                 logreporte.save(usuario_id=eUser.id)
#
#         else:
#             """SE MANTIENE PARA VERSIONES ANTERIORES"""
#             if 'variableqr' in params:
#                 variable = params['variableqr']
#                 pdfname = eReporte.nombre + variable
#                 matricula_id = None
#                 inscripcion_id = None
#                 parametros_reporte = eReporte.parametroreporte_set.get(nombre='variableqr')
#                 if (parametros_reporte.extra).upper() == "MATRICULA":
#                     matricula_id = variable
#                     matricula = Matricula.objects.get(pk=matricula_id)
#                     inscripcion = matricula.inscripcion
#                     periodo = matricula.nivel.periodo
#                     confirmar_automatricula_admision = inscripcion.tiene_automatriculaadmision_por_confirmar(periodo)
#                     if confirmar_automatricula_admision and periodo.limite_agregacion < datetime.now().date():
#                         cordinacionid = inscripcion.carrera.coordinacion_carrera().id
#                         if cordinacionid in [9]:
#                             return False, u"Estimado aspirante, el período de confirmación de la automatrícula ha culminado, usted no se encuentra matriculado", {}
#                 else:
#                     inscripcion_id = variable
#
#                 itemdescarga = ReporteDescarga(reporte=eReporte, matricula_id=matricula_id, inscripcion_id=encrypt(inscripcion_id))
#                 itemdescarga.save(usuario_id=eUser.id)
#                 cambiaruta = 1
#         folder = os.path.join(os.path.join(SITE_STORAGE, 'media', 'documentos', 'userreports', remove_accents(eUser.username), ''))
#         # folder = '/mnt/nfs/home/storage/media/qrcode/evaluaciondocente/'
#         rutapdf = folder + pdfname + '.pdf'
#         if os.path.isfile(rutapdf):
#             os.remove(rutapdf)
#
#         runjrcommand = [JR_JAVA_COMMAND, '-jar',
#                         os.path.join(JR_RUN, 'jasperstarter.jar'),
#                         'pr', eReporte.archivo.file.name,
#                         '--jdbc-dir', JR_RUN,
#                         '-f', tipo,
#                         '-t', 'postgres',
#                         '-H', DATABASES['sga_select']['HOST'],
#                         '-n', DATABASES['sga_select']['NAME'],
#                         '-u', DATABASES['sga_select']['USER'],
#                         '-p', f"'{DATABASES['sga_select']['PASSWORD']}'",
#                         '--db-port', DATABASES['sga_select']['PORT'],
#                         '-o', output_folder + os.sep + pdfname]
#         parametros = eReporte.parametros()
#         paramlist = [transform_jasperstarter(p, params) for p in parametros]
#         if paramlist:
#             runjrcommand.append('-P')
#             for parm in paramlist:
#                 if 'qr=' in parm and 'true' in parm:
#                     if eReporte.version == 2:
#                         url = pyqrcode.create(base_url + "/".join([MEDIA_URL, 'documentos', 'userreports', remove_accents(eUser.username), pdfname + "." + tipo]))
#                         url.png(os.path.join(os.path.join(SITE_STORAGE, 'media', 'qrcode', pdfname + '.png', 10, '#000000')))
#                     else:
#                         url = pyqrcode.create("http://sga.unemi.edu.ec" + "/".join([MEDIA_URL, 'documentos', 'userreports', remove_accents(eUser.username), pdfname + "." + tipo]))
#                         url.png('/var/lib/django/sistemagestion/media/qrcode/' + pdfname + '.png', 10, '#000000')
#                     runjrcommand.append(u'qr=' + pdfname + '.png')
#                 else:
#                     runjrcommand.append(parm)
#         else:
#             runjrcommand.append('-P')
#
#         runjrcommand.append(u'userweb=' + str(eUser.username))
#         if eReporte.version == 2:
#             runjrcommand.append(u'MEDIA_DIR=' + str("/".join([MEDIA_ROOT, ''])))
#             if DEBUG:
#                 runjrcommand.append(u'IMAGE_DIR=' + str("/".join([SUBREPOTRS_FOLDER, ''])))
#             else:
#                 runjrcommand.append(u'IMAGE_DIR=' + str(SUBREPOTRS_FOLDER))
#             runjrcommand.append(u'SUBREPORT_DIR=' + eReporte.ruta_subreport())
#             if isQR:
#                 runjrcommand.append(u'URL_QR=' + str(base_url + "/".join([MEDIA_URL, 'documentos', 'userreports', remove_accents(eUser.username), pdfname + "." + tipo])))
#                 runjrcommand.append(u'CODIGO_QR=' + str(codigo))
#                 runjrcommand.append(u'CERTIFICADO_ID=' + str(certificado.id))
#         else:
#             runjrcommand.append(u'SUBREPORT_DIR=' + str(SUBREPOTRS_FOLDER))
#         mens = ''
#         mensaje = ''
#         for m in runjrcommand:
#             mens += ' ' + m
#         urlbase = get_variable('SITE_URL_SGA')
#         reportfile = "/".join([MEDIA_URL, 'documentos', 'userreports', remove_accents(eUser.username), pdfname + "." + tipo])
#         # reportfile = f"{urlbase}{reportfile}"
#         if eReporte.es_background:
#             data['logreporte'] = logreporte
#             ReportBackground(params=params, data=data, reporte=eReporte, mensaje=mens, reportfile=reportfile, certificado=certificado, solicitud=eSolicitud).start()
#             lista_correos = []
#             if 'dirigidos' in params:
#                 dirigidos = json.loads(params['dirigidos'])
#                 personas = Persona.objects.filter(pk__in=dirigidos)
#                 for persona in personas:
#                     lista_correos.extend(persona.lista_emails())
#             if not 'no_persona_session' in params:
#                 if ePersona:
#                     lista_correos.extend(ePersona.lista_emails())
#             mensaje = 'El reporte se está realizando. Verifique los correos: %s después de unos minutos.' % (", ".join(lista_correos))
#         else:
#             if DEBUG:
#                 runjr = subprocess.run(mens, shell=True, check=True)
#                 # print('codigo:', mens)
#             else:
#                 runjr = subprocess.call(mens.encode("latin1"), shell=True)
#             if certificado:
#                 if certificado.funcionadjuntar:
#                     if certificado.funcionadjuntar.get_nombrefuncion_display() == 'adjuntar_malla_curricular':
#                         result, mensaje, reportfile_aux = certificado.funcionadjuntar.adjuntar_malla_curricular(logreporte)
#                         if not result:
#                             raise NameError(mensaje)
#                         reportfile = "/".join([MEDIA_URL, 'documentos', 'userreports', remove_accents(eUser.username), reportfile_aux])
#             reportfile = f"{urlbase}{reportfile}"
#             if eReporte.enviar_email:
#                 ids_personas = []
#                 if 'dirigidos' in params:
#                     dirigidos = json.loads(params['dirigidos'])
#                     personas = Persona.objects.filter(pk__in=dirigidos)
#                     for persona in personas:
#                         ids_personas.append(persona.id)
#                     if not 'no_persona_session' in params:
#                         if ePersona:
#                             ids_personas.append(ePersona.id)
#                 # reportfile = "/".join([MEDIA_URL, 'documentos', 'userreports', remove_accents(eUser.username), pdfname + "." + tipo])
#                 # reportfile = f"{urlbase}{reportfile}"
#                 for persona in Persona.objects.filter(pk__in=ids_personas).distinct():
#                     send_html_mail(eReporte.descripcion,
#                                    "reportes/emails/reporte_generacion.html",
#                                    {'sistema': u'SGA+',
#                                     'persona': persona,
#                                     'reporte': eReporte,
#                                     'reportfile': reportfile,
#                                     't': miinstitucion()},
#                                    persona.lista_emails(),
#                                    [],
#                                    cuenta=variable_valor('CUENTAS_CORREOS')[0]
#                                    )
#             if eSolicitud:
#                 if os.path.isfile(rutapdf):
#                     with open(rutapdf, 'rb') as f:
#                         archivo = f.read()
#                     buffer = io.BytesIO()
#                     buffer.write(archivo)
#                     pdf = buffer.getvalue()
#                     buffer.seek(0)
#                     buffer.close()
#                     fecha = datetime.now().date()
#                     hora = datetime.now().time()
#                     nombre_archivo_resultado = f"resultado_{fecha.year.__str__() + fecha.month.__str__() + fecha.day.__str__()}_{hora.hour.__str__() + hora.minute.__str__() + hora.second.__str__()}.pdf"
#                     eSolicitud.archivo_respuesta.save(nombre_archivo_resultado, ContentFile(pdf))
#                 eSolicitud.en_proceso = False
#
#                 if eSolicitud.perfil.inscripcion.carrera.coordinacion_carrera().id == 7 and eSolicitud.origen_object_id in [57, 39, 12]:
#                     eSolicitud.estado = 22
#                     eSolicitud.save()
#                 else:
#                     eSolicitud.estado = 2
#                     eSolicitud.save()
#                 observacion = f'Cambio de estado por proceso de entrega'
#                 #if ePersona.es_estudiante():
#                 #    observacion= f'Cambio de estado por generación'
#                 eHistorialSolicitud = HistorialSolicitud(solicitud=eSolicitud,
#                                                          observacion=observacion,
#                                                          fecha=datetime.now().date(),
#                                                          hora=datetime.now().time(),
#                                                          estado=eSolicitud.estado,
#                                                          responsable=ePersona,
#                                                          #archivo=eSolicitud.archivo_respuesta,
#                                                          )
#                 eHistorialSolicitud.save()
#                 titulo = f'Cambio de estado de la solicitud código {eSolicitud.codigo}'
#                 cuerpo = f'Solicitud código {eSolicitud.codigo} se encuentra en estado {eSolicitud.get_estado_display()}'
#                 eNotificacion = Notificacion(titulo=titulo,
#                                              cuerpo=cuerpo,
#                                              destinatario=eSolicitud.perfil.persona,
#                                              perfil=eSolicitud.perfil,
#                                              url='/alu_secretaria/mis_pedidos',
#                                              prioridad=1,
#                                              app_label='SIE',
#                                              fecha_hora_visible=datetime.now() + timedelta(days=2),
#                                              tipo=1,
#                                              en_proceso=False,
#                                              content_type=ContentType.objects.get_for_model(eSolicitud),
#                                              object_id=eSolicitud.pk,
#                                              )
#                 eNotificacion.save()
#
#         # sp = os.path.split(eReporte.archivo.file.name)
#         # return ok_json({'reportfile': "/".join([MEDIA_URL, 'documentos', 'userreports', remove_accents(eUser.username), pdfname + "." + tipo])})
#         aData = {'es_background': eReporte.es_background,
#                  'reportfile': reportfile
#                  }
#         return True, aData, mensaje
#     else:
#         tipo = params['rt']
#         d = datetime.now()
#         output_folder = os.path.join(JR_USEROUTPUT_FOLDER, remove_accents(eUser.username))
#         try:
#             os.makedirs(output_folder)
#         except Exception as ex:
#             pass
#         mensaje = ''
#
#         if tipo == 'pdf':
#             nombre = eReporte.nombre + d.strftime('%Y%m%d_%H%M%S') + ".pdf"
#             paramlist = {}
#             for p in eReporte.parametros():
#                 paramlist.update({p.nombre: fixparametro(p.tipo, params[p.nombre])})
#             global django_engine
#             django_engine = engines['django']
#             d = locals()
#             exec(eReporte.vista, globals(), d)
#             reportefinal = d['vistareporte'](eReporte, parametros=paramlist)
#             filepdf = open(output_folder + os.sep + nombre, "w+b")
#             pdf = pisa.pisaDocument(StringIO.BytesIO(reportefinal), dest=filepdf, link_callback=fetch_resources)
#             filepdf.close()
#         else:
#             nombre = eReporte.nombre + d.strftime('%Y%m%d_%H%M%S') + ".xls"
#             paramlist = {}
#             for p in eReporte.parametros():
#                 paramlist.update({p.nombre: fixparametro(p.tipo, params[p.nombre])})
#             d = locals()
#             exec(eReporte.vista, globals(), d)
#             book = d['vistareporte'](eReporte, parametros=paramlist)
#             output_folder = os.path.join(JR_USEROUTPUT_FOLDER, remove_accents(eUser.username))
#             filename = os.path.join(output_folder, nombre)
#             book.save(filename)
#         reportfile = "/".join([MEDIA_URL, 'documentos', 'userreports', remove_accents(eUser.username), nombre])
#         urlbase = get_variable('SITE_URL_SGA')
#         reportfile = f"{urlbase}{reportfile}"
#         aData = {'es_background': eReporte.es_background,
#                  'reportfile': reportfile
#                  }
#
# try:
#     eInscripciones  = Inscripcion.objects.filter(status=True, id__in=[58188])
#     for eInscripcion in eInscripciones:
#         ePerfilUsuario = PerfilUsuario.objects.get(status=True, inscripcion=eInscripcion)
#         eCoordinacion = eInscripcion.carrera.coordinacion_set.all()[0]
#         ePersona = ePerfilUsuario.persona
#         idencrypt = encrypt(eInscripcion.id)
#         parametros = {'n': 'rpt_certificado_egresado_grado', 'vqr': idencrypt, 'rt': 'pdf'}
#         eCertificado = Certificado.objects.get(reporte__id=667)
#         eServicio = eCertificado.servicio
#         valor = Decimal(eCertificado.costo).quantize(Decimal('.01'))
#         eContentTypeCertificado = ContentType.objects.get_for_model(eCertificado)
#         eSolicitudes = Solicitud.objects.filter(Q(en_proceso=True) | Q(estado__in=[1]), status=True,
#                                                 perfil_id=ePerfilUsuario.pk, servicio_id=eServicio.pk,
#                                                 origen_content_type_id=eContentTypeCertificado.pk,
#                                                 origen_object_id=eCertificado.id).exclude(servicio__proceso=8)
#         if not eSolicitudes.values("id").exists():
#             result = generar_codigo_solicitud(eServicio)
#             success = result.get('success', False)
#             codigo = result.get('codigo', None)
#             secuencia = result.get('secuencia', 0)
#             suffix = result.get('suffix', None)
#             prefix = result.get('prefix', None)
#             if not success:
#                 raise NameError(u"Código de solicitud no se pudo generar")
#             if codigo is None:
#                 raise NameError(u"Código de solicitud no se pudo generar")
#             if secuencia == 0:
#                 raise NameError(u"Secuenia del código de solicitud no se pudo generar")
#             if suffix is None:
#                 raise NameError(u"Sufijo del código de solicitud no se pudo generar")
#             if prefix is None:
#                 raise NameError(u"Prefijo del código de solicitud no se pudo generar")
#
#             eSolicitud = Solicitud(codigo=codigo,
#                                    secuencia=secuencia,
#                                    prefix=prefix,
#                                    suffix=suffix,
#                                    perfil_id=ePerfilUsuario.pk,
#                                    servicio_id=eServicio.pk,
#                                    origen_content_type=eContentTypeCertificado,
#                                    origen_object_id=eCertificado.pk,
#                                    descripcion=f'Solicitud ({codigo}) del Certificado académico con código {eCertificado.codigo} - {eCertificado.certificacion}',
#                                    fecha=datetime.now().date(),
#                                    hora=datetime.now().time(),
#                                    estado=22,
#                                    cantidad=1,
#                                    valor_unitario=valor,
#                                    subtotal=valor,
#                                    iva=0,
#                                    descuento=0,
#                                    en_proceso=False,
#                                    parametros=parametros,
#                                    tiempo_cobro=eCertificado.tiempo_cobro)
#             eSolicitud.save()
#             eHistorialSolicitud = HistorialSolicitud(solicitud=eSolicitud,
#                                                      observacion=eSolicitud.descripcion,
#                                                      fecha=eSolicitud.fecha,
#                                                      hora=eSolicitud.hora,
#                                                      estado=eSolicitud.estado,
#                                                      responsable=eSolicitud.perfil.persona,
#                                                      tipodocumento=3)
#             eHistorialSolicitud.save()
#
#             parametros = eSolicitud.parametros
#             ids_persona = [x.id for x in Persona.objects.filter(pk__in=[eSolicitud.perfil.persona_id])]
#             parametros['dirigidos'] = json.dumps(ids_persona)
#             parametros['no_persona_session'] = True
#             parametros['app'] = 'sie'
#
#             generar_certificado_digital(eSolicitud.perfil.persona, eCertificado.reporte, eSolicitud.perfil.persona.usuario, parametros, eSolicitud)
#             eSolicitud.estado=22
#             eSolicitud.save()
#             print(f'Certificado generado! - {eSolicitud.perfil.inscripcion.persona}')
# except Exception as ex:
#     print(ex)
#     print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))



# print("ZIP PLANTILLA")
# try:
#     directory = os.path.join(MEDIA_ROOT, 'reportes', 'zip', 'plantilla_unemi1.zip')
#     plantillas = DistributivoPersona.objects.filter(status=True).order_by('id')
#     url = os.path.join(SITE_STORAGE, 'media', 'zip', 'plantilla_unemi1.zip')
#     url_zip = url
#     fantasy_zip = zipfile.ZipFile(url, 'w')
#     c = 0
#     for plantilla in plantillas:
#         if plantilla.persona.foto():
#             url_archivo = (SITE_STORAGE + plantilla.persona.foto().foto.url).replace('\\', '/')
#             _name = str(plantilla.id)
#             folder = os.path.join(SITE_STORAGE, 'media', 'fotosplantilla1', '')
#             if not os.path.exists(folder):
#                 os.makedirs(folder)
#             url_file_generado = f'{_name}.jpg'
#             ruta_creacion = SITE_STORAGE
#             ruta_creacion = ruta_creacion.replace('\\', '/')
#             shutil.copy(url_archivo, ruta_creacion + '/media/fotosplantilla1/' + url_file_generado)
#             c += 1
#             print(f'{c} - {plantilla.id} - {plantilla.persona}')
#             fantasy_zip.write(ruta_creacion + '/media/fotosplantilla1/' + url_file_generado,
#                               os.path.basename(ruta_creacion + '/media/fotosplantilla1/' + url_file_generado))
#
#     print('Finaliza')
#     #     c += 1
#     #     print(f'{c} - {plantilla.persona}')
#     fantasy_zip.close()
#     response = HttpResponse(directory, content_type='application/zip')
#     response['Content-Disposition'] = 'attachment; filename=plantilla_unemi1.zip'
#     print(directory)
# except Exception as ex:
#     pass

# try:
#     cohortes = CohorteMaestria.objects.filter(status=True, procesoabierto=False, maestriaadmision__carrera__id=236).exclude(pk=192)
#     for cohorte in cohortes:
#         if CohorteMaestria.objects.filter(status=True, maestriaadmision__carrera__id=267, descripcion=cohorte.descripcion, procesoabierto=False).exists():
#             cohortenueva = CohorteMaestria.objects.get(status=True, maestriaadmision__carrera__id=267, descripcion=cohorte.descripcion, procesoabierto=False)
#             postulantes = InscripcionCohorte.objects.filter(status=True, cohortes=cohorte,
#                                                             cohortes__maestriaadmision__carrera__id=236,
#                                                             itinerario=1).exclude(cohortes__id__in=[192, 206, 207]).order_by('id')
#             c = 0
#             for postulante in postulantes:
#                 postulante.cohortes = cohortenueva
#                 postulante.save()
#
#                 requisitos = RequisitosMaestria.objects.filter(status=True, cohorte=cohortenueva)
#
#                 for reqma in requisitos:
#                     if EvidenciaRequisitosAspirante.objects.filter(inscripcioncohorte=postulante, requisitos__requisito=reqma.requisito).exists():
#                         evi = EvidenciaRequisitosAspirante.objects.filter(inscripcioncohorte=postulante, requisitos__requisito=reqma.requisito).first()
#                         evi.requisitos = reqma
#                         evi.save()
#
#                 if Rubro.objects.filter(status=True, inscripcion=postulante).exists():
#                     rubros = Rubro.objects.filter(status=True, inscripcion=postulante)
#                     for rubro in rubros:
#                         rubro.cohortemaestria = cohortenueva
#                         rubro.tipo = cohorte.tiporubro
#                         rubro.save()
#                 if postulante.Configfinanciamientocohorte:
#                     if ConfigFinanciamientoCohorte.objects.filter(descripcion=postulante.Configfinanciamientocohorte.descripcion, cohorte=cohortenueva).exists():
#                         finan = ConfigFinanciamientoCohorte.objects.filter(descripcion=postulante.Configfinanciamientocohorte.descripcion, cohorte=cohortenueva).first()
#                         postulante.Configfinanciamientocohorte = finan
#                         postulante.save()
#                 c += 1
#                 print(f'{c}/{postulantes.count()} - Cedula: {postulante.inscripcionaspirante.persona.cedula} - {postulante}')
#
#     print("CIENCIAS SOCIALES")
#     cohortes = CohorteMaestria.objects.filter(status=True, procesoabierto=False, maestriaadmision__carrera__id=236).exclude(pk=192)
#     for cohorte in cohortes:
#         if CohorteMaestria.objects.filter(status=True, maestriaadmision__carrera__id=268, descripcion=cohorte.descripcion, procesoabierto=False).exists():
#             cohortenueva = CohorteMaestria.objects.get(status=True, maestriaadmision__carrera__id=268, descripcion=cohorte.descripcion, procesoabierto=False)
#             postulantes = InscripcionCohorte.objects.filter(status=True, cohortes=cohorte,
#                                                             cohortes__maestriaadmision__carrera__id=236,
#                                                             itinerario=3).exclude(cohortes__id__in=[192, 206, 207]).order_by('id')
#             c = 0
#             for postulante in postulantes:
#                 postulante.cohortes = cohortenueva
#                 postulante.save()
#
#                 requisitos = RequisitosMaestria.objects.filter(status=True, cohorte=cohortenueva)
#
#                 for reqma in requisitos:
#                     if EvidenciaRequisitosAspirante.objects.filter(inscripcioncohorte=postulante, requisitos__requisito=reqma.requisito).exists():
#                         evi = EvidenciaRequisitosAspirante.objects.filter(inscripcioncohorte=postulante, requisitos__requisito=reqma.requisito).first()
#                         evi.requisitos = reqma
#                         evi.save()
#
#                 if Rubro.objects.filter(status=True, inscripcion=postulante).exists():
#                     rubros = Rubro.objects.filter(status=True, inscripcion=postulante)
#                     for rubro in rubros:
#                         rubro.cohortemaestria = cohortenueva
#                         rubro.tipo = cohorte.tiporubro
#                         rubro.save()
#                 if postulante.Configfinanciamientocohorte:
#                     if ConfigFinanciamientoCohorte.objects.filter(descripcion=postulante.Configfinanciamientocohorte.descripcion, cohorte=cohortenueva).exists():
#                         finan = ConfigFinanciamientoCohorte.objects.filter(descripcion=postulante.Configfinanciamientocohorte.descripcion, cohorte=cohortenueva).first()
#                         postulante.Configfinanciamientocohorte = finan
#                         postulante.save()
#                 c += 1
#                 print(f'{c}/{postulantes.count()} - Cedula: {postulante.inscripcionaspirante.persona.cedula} - {postulante}')
# except Exception as ex:
#     pass
# print(u"Inicio 2")
#
# def titulo_4_nivel(persona):
#     try:
#         titulo = ''
#         if Titulacion.objects.filter(status=True, persona=persona, titulo__nivel__id__in=[4, 21, 23, 30]).exists():
#             titulos = Titulacion.objects.filter(status=True, persona=persona, titulo__nivel__id__in=[4, 21, 23, 30]).order_by('-id')
#             for idx, t in enumerate(titulos):
#                 titulo += t.titulo.nombre
#                 if idx < len(titulos) - 1:
#                     titulo += ', '
#         return titulo
#     except Exception as ex:
#         pass
#
# def universidad_titulo_4_nivel(persona):
#     try:
#         titulo = ''
#         if Titulacion.objects.filter(status=True, persona=persona, titulo__nivel__id__in=[4, 21, 23, 30]).exists():
#             titulos = Titulacion.objects.filter(status=True, persona=persona, titulo__nivel__id__in=[4, 21, 23, 30]).order_by('-id')
#             for idx, t in enumerate(titulos):
#                 if t.institucion:
#                     titulo += t.institucion.nombre
#                 else:
#                     titulo += 'NO REGISTRA'
#                 if idx < len(titulos) - 1:
#                     titulo += ', '
#         return titulo
#     except Exception as ex:
#         pass
#
# def trabaja_en(persona):
#     try:
#         trabajo = 'NO REGISTRA'
#         if PersonaSituacionLaboral.objects.filter(status=True, persona=persona).exists():
#             trabajo = PersonaSituacionLaboral.objects.filter(status=True, persona=persona).first()
#             if trabajo.lugartrabajo:
#                 trabajo = trabajo.lugartrabajo
#             elif trabajo.negocio:
#                 trabajo = trabajo.negocio
#         return trabajo
#     except Exception as ex:
#         pass
#
#
# def imprimirreporte():
#     try:
#         estado = ''
#
#         __author__ = 'Unemi'
#
#         output = io.BytesIO()
#         name_document = 'reporte_graduados_grado'
#         nombre_archivo = name_document + "_9.xlsx"
#         directory = os.path.join(MEDIA_ROOT, 'reportes', 'gestion', nombre_archivo)
#         workbook = xlsxwriter.Workbook(directory, {'constant_memory': True})
#         ws = workbook.add_worksheet('resultados')
#         ws.set_column(0, 0, 10)
#         ws.set_column(1, 1, 15)
#         ws.set_column(2, 2, 15)
#         ws.set_column(3, 3, 35)
#         ws.set_column(4, 4, 25)
#         ws.set_column(5, 5, 40)
#         ws.set_column(6, 6, 40)
#         ws.set_column(7, 7, 40)
#         ws.set_column(8, 8, 15)
#         ws.set_column(9, 9, 15)
#         ws.set_column(10, 10, 45)
#         ws.set_column(11, 11, 35)
#         ws.set_column(12, 12, 30)
#         ws.set_column(13, 13, 40)
#         ws.set_column(14, 14, 35)
#
#         formatotitulo_filtros = workbook.add_format(
#             {'bold': 1, 'text_wrap': True, 'border': 1, 'align': 'center', 'font_size': 14})
#
#         formatoceldacab = workbook.add_format(
#             {'align': 'center', 'bold': 1, 'border': 1, 'text_wrap': True, 'fg_color': '#1C3247', 'font_color': 'white'})
#         formatoceldaleft = workbook.add_format(
#             {'text_wrap': True, 'align': 'center', 'valign': 'vcenter', 'border': 1})
#
#         formatoceldaleft2 = workbook.add_format(
#             {'num_format': '$ #,##0.00', 'text_wrap': True, 'align': 'center', 'valign': 'vcenter', 'border': 1, 'bold': 1})
#
#         formatoceldaleft3 = workbook.add_format(
#             {'text_wrap': True, 'align': 'right', 'valign': 'vcenter', 'border': 1, 'bold': 1})
#
#         decimalformat = workbook.add_format(
#             {'num_format': '$ #,##0.00', 'text_wrap': True, 'align': 'center', 'valign': 'vcenter', 'border': 1})
#
#         decimalformat2 = workbook.add_format(
#             {'num_format': '$ #,##0.00', 'text_wrap': True, 'align': 'center', 'valign': 'vcenter', 'border': 1, 'bold': 1})
#
#         ws.write(2, 0, 'N°', formatoceldacab)
#         ws.write(2, 1, 'FECHA GRADUACIÓN', formatoceldacab)
#         ws.write(2, 2, 'CÉDULA', formatoceldacab)
#         ws.write(2, 3, 'CARRERA', formatoceldacab)
#         ws.write(2, 4, 'MODALIDAD', formatoceldacab)
#         ws.write(2, 5, 'NOMBRES', formatoceldacab)
#         ws.write(2, 6, 'EMAIL PERSONAL', formatoceldacab)
#         ws.write(2, 7, 'EMAIL CORPORATIVO- INSTITUCIONAL', formatoceldacab)
#         ws.write(2, 8, 'TELÉFONO 1', formatoceldacab)
#         ws.write(2, 9, 'TELÉFONO 2', formatoceldacab)
#         ws.write(2, 10, 'TITULO TERCER NIVEL OBTENIDO', formatoceldacab)
#         ws.write(2, 11, 'INSTITUCION DONDE OBTUVO EL 3ER NIVEL', formatoceldacab)
#         ws.write(2, 12, 'TITULO  CUARTO NIVEL OBTENIDO', formatoceldacab)
#         ws.write(2, 13, 'INSTITUCION DONDE OBTUVO EL 4TO NIVEL', formatoceldacab)
#         ws.write(2, 14, 'TRABAJA EN', formatoceldacab)
#
#         graduados = Graduado.objects.filter(status=True, fechagraduado__year__gte=2013).exclude(inscripcion__carrera__coordinacion__id=7).order_by('inscripcion__carrera__nombre')
#
#         filas_recorridas = 4
#         cont = 1
#
#         for graduado in graduados:
#             titulo = ''
#             if graduado.inscripcion.carrera.titulootorga:
#                 titulo = graduado.inscripcion.carrera.titulootorga
#             elif graduado.inscripcion.carrera.tituloobtenido:
#                 titulo = graduado.inscripcion.carrera.tituloobtenido
#             elif graduado.nombretitulo:
#                 titulo = graduado.nombretitulo
#             elif graduado.inscripcion.carrera.tituloobtenidohombre:
#                 titulo = graduado.inscripcion.carrera.tituloobtenidohombre
#             elif graduado.inscripcion.carrera.tituloobtenidomujer:
#                 titulo = graduado.inscripcion.carrera.tituloobtenidomujer
#
#             ws.write('A%s' % filas_recorridas, str(cont), formatoceldaleft)
#             ws.write('B%s' % filas_recorridas, str(graduado.fechagraduado), formatoceldaleft)
#             ws.write('C%s' % filas_recorridas, str(graduado.inscripcion.persona.cedula), formatoceldaleft)
#             ws.write('D%s' % filas_recorridas, str(graduado.inscripcion.carrera), formatoceldaleft)
#             ws.write('E%s' % filas_recorridas, str(graduado.inscripcion.carrera.get_modalidad_display()), formatoceldaleft)
#             ws.write('F%s' % filas_recorridas, str(graduado.inscripcion.persona.nombre_completo_inverso()), formatoceldaleft)
#             ws.write('G%s' % filas_recorridas, str(graduado.inscripcion.persona.email if graduado.inscripcion.persona.email else 'NO REGISTRA'), formatoceldaleft)
#             ws.write('H%s' % filas_recorridas, str(graduado.inscripcion.persona.emailinst if graduado.inscripcion.persona.email else 'NO REGISTRA'), formatoceldaleft)
#             ws.write('I%s' % filas_recorridas, str(graduado.inscripcion.persona.telefono), formatoceldaleft)
#             ws.write('J%s' % filas_recorridas, str(graduado.inscripcion.persona.telefono2), formatoceldaleft)
#             ws.write('K%s' % filas_recorridas, str(titulo), formatoceldaleft)
#             ws.write('L%s' % filas_recorridas, str('UNIVERSIDAD ESTATAL DE MILAGRO'), formatoceldaleft)
#             ws.write('M%s' % filas_recorridas, str(titulo_4_nivel(graduado.inscripcion.persona)), formatoceldaleft)
#             ws.write('N%s' % filas_recorridas, str(universidad_titulo_4_nivel(graduado.inscripcion.persona)), formatoceldaleft)
#             ws.write('O%s' % filas_recorridas, str(trabaja_en(graduado.inscripcion.persona)), formatoceldaleft)
#
#             filas_recorridas += 1
#
#             print(f'Persona: {graduado.inscripcion.persona} - Carrera: {graduado.inscripcion.carrera} - {cont} / {graduados.count()}')
#             cont += 1
#
#         workbook.close()
#         # output.seek(0)
#         # fecha_hora_actual = datetime.now().date()
#         # filename = 'Listado_ventas_' + str(fecha_hora_actual) + '.xlsx'
#         response = HttpResponse(directory, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#         response['Content-Disposition'] = 'attachment; filename=%s' % name_document
#         url_file = "{}reportes/gestion/{}".format(MEDIA_URL, nombre_archivo)
#         print(url_file)
#     except Exception as ex:
#         pass
#
# imprimirreporte()


# print(u"********************Migración bachillerato***************************")
# print(u"********************DERECHO INSCIRTOS PENAL 14-15-16**************************")
# try:
#     #Excluir a los leads de la maestría vieja
#     # idins = InscripcionCohorte.objects.filter(status=True, cohortes__id=192, inscripcion__isnull=False).values_list('inscripcion__id', flat=True)
#     inscripciones = Inscripcion.objects.filter(status=True, id__in=[326349,
# 326362,
# 326365,
# 326829,
# 327176,
# 327201,
# 326368,
# 326788,
# 326845,
# 326853,
# 326871,
# 326377,
# 326379,
# 326378,
# 326546,
# 326612,
# 326615,
# 327069,
# 326806,
# 327072,
# 326815,
# 326816,
# 326950,
# 326960,
# 326967,
# 326645,
# 326651,
# 326988,
# 327009,
# 327021,
# 327024,
# 327026,
# 326732,
# 326750,
# 326760,
# 326773,
# 253348,
# 327094,
# 327102,
# 327107,
# 327167,
# 326121,
# 326346])
#     #nueva carrera
#     carrera = Carrera.objects.get(status=True, pk=266)
#     malla = Malla.objects.get(status=True, pk=500)
#     cont = 0
#     for inscripcion in inscripciones:
#         inscripcion.carrera = carrera
#         inscripcion.save()
#
#         for eInscripcionMalla in inscripcion.inscripcionmalla_set.all():
#             eInscripcionMalla.status = False
#             eInscripcionMalla.save()
#         eInscripcionMallas = inscripcion.inscripcionmalla_set.filter(inscripcion=inscripcion,
#                                                                      malla=malla)
#         if eInscripcionMallas.values("id").exists():
#             eInscripcionMalla = eInscripcionMallas[0]
#             eInscripcionMalla.status = True
#         else:
#             eInscripcionMalla = InscripcionMalla(inscripcion=inscripcion,
#                                                  malla=malla)
#         eInscripcionMalla.save()
#         inscripcion.actualizar_creditos()
#         inscripcion.actualizar_nivel()
#
#         cont += 1
#
#         print(f'{cont} - Cédula: {inscripcion.persona.cedula} - Inscrito: {inscripcion.persona} - Carrera: {inscripcion.carrera.id} - {cont}/{inscripciones.count()}')
#
#     # print("************************+CIENCIAS SOCIALES**********************************")
#     # #Excluir a los leads de la maestría vieja
#     # idins = InscripcionCohorte.objects.filter(status=True, cohortes__id=192, inscripcion__isnull=False).values_list('inscripcion__id', flat=True)
#     # inscripciones = Inscripcion.objects.filter(status=True, carrera__id=236, itinerario=3)
#     # #nueva carrera
#     # carrera = Carrera.objects.get(status=True, pk=268)
#     # malla = Malla.objects.get(status=True, pk=502)
#     # cont = 0
#     # for inscripcion in inscripciones:
#     #     inscripcion.carrera = carrera
#     #     inscripcion.save()
#     #
#     #     for eInscripcionMalla in inscripcion.inscripcionmalla_set.all():
#     #         eInscripcionMalla.status = False
#     #         eInscripcionMalla.save()
#     #     eInscripcionMallas = inscripcion.inscripcionmalla_set.filter(inscripcion=inscripcion,
#     #                                                                  malla=malla)
#     #     if eInscripcionMallas.values("id").exists():
#     #         eInscripcionMalla = eInscripcionMallas[0]
#     #         eInscripcionMalla.status = True
#     #     else:
#     #         eInscripcionMalla = InscripcionMalla(inscripcion=inscripcion,
#     #                                              malla=malla)
#     #     eInscripcionMalla.save()
#     #     inscripcion.actualizar_creditos()
#     #     inscripcion.actualizar_nivel()
#     #
#     #     cont += 1
#     #
#     #     print(f'{cont} - Cédula: {inscripcion.persona.cedula} - Inscrito: {inscripcion.persona} - Carrera: {inscripcion.carrera.id} - {inscripcion.itinerario} - {cont}/{inscripciones.count()}')
# except Exception as ex:
#     pass
# try:
#     a=0
#     inscritos = Inscripcion.objects.filter(id__in=[222758, 250700, 222757]).distinct()
#     consti = Carrera.objects.get(pk=215)
#     penal = Carrera.objects.get(pk=266)
#     for inscrito in inscritos:
#         if inscrito.itinerario == 1:
#             postulante =InscripcionCohorte.objects.filter(status=True, cohortes__maestriaadmision__carrera__id__in=[215,266],
#                                                           inscripcionaspirante__persona=inscrito.persona).order_by('-id').first()
#             print(f'Inscrito: {postulante} - cedula: {inscrito.persona.cedula}')
#             # matricula = Matricula.objects.filter(status=True, inscripcion=inscrito).exclude(nivel__periodo__nombre__icontains='TITULA').order_by('-id').first()
#
#             inscripcion = Inscripcion(persona=postulante.inscripcionaspirante.persona,
#                                       fecha=datetime.now(),
#                                       carrera=penal,
#                                       modalidad=inscrito.modalidad,
#                                       sesion=inscrito.sesion,
#                                       sede=inscrito.sede,
#                                       colegio='',
#                                       aplica_b2=True,
#                                       fechainicioprimernivel=datetime.now(),
#                                       fechainiciocarrera=datetime.now(),
#                                       itinerario=2)
#             inscripcion.save()
#
#             print(f'Nueva inscripcion: {inscripcion} - Carrera: {inscripcion.carrera.id}')
#
#             postulante.inscripcion = inscripcion
#             postulante.save()
#             print(f'Postulación actualizada')
#             # matricula.inscripcion = inscripcion
#             # matricula.save()
#             # print(f'Matricula actualizada')
#
#             postulante.inscripcionaspirante.persona.crear_perfil(inscripcion=inscripcion)
#             a += 1
#             print(f'{a}/{inscritos.count()}')
#             if PerfilUsuario.objects.values_list('id').filter(inscripcion=inscripcion, status=True).exists():
#                 perfil = PerfilUsuario.objects.filter(inscripcion=inscripcion, status=True).last()
#                 perfil.inscripcionprincipal = True
#                 perfil.save()
#
#                 documentos = DocumentosDeInscripcion(inscripcion=inscripcion,
#                                                      titulo=False,
#                                                      acta=False,
#                                                      cedula=False,
#                                                      votacion=False,
#                                                      actaconv=False,
#                                                      partida_nac=False,
#                                                      pre=False,
#                                                      observaciones_pre='',
#                                                      fotos=False)
#                 documentos.save()
#                 preguntasinscripcion = inscripcion.preguntas_inscripcion()
#                 inscripciontesdrive = InscripcionTesDrive(inscripcion=inscripcion,
#                                                           licencia=False,
#                                                           record=False,
#                                                           certificado_tipo_sangre=False,
#                                                           prueba_psicosensometrica=False,
#                                                           certificado_estudios=False)
#                 inscripciontesdrive.save()
#                 inscripcion.malla_inscripcion()
#                 inscripcion.actualizar_nivel()
#                 if USA_TIPOS_INSCRIPCIONES:
#                     inscripciontipoinscripcion = InscripcionTipoInscripcion(inscripcion=inscripcion,
#                                                                             tipoinscripcion_id=TIPO_INSCRIPCION_INICIAL)
#                     inscripciontipoinscripcion.save()
#
#         elif inscrito.itinerario == 2:
#             postulante =InscripcionCohorte.objects.filter(status=True, cohortes__maestriaadmision__carrera__id__in=[215,266],
#                                                           inscripcionaspirante__persona=inscrito.persona).order_by('-id').first()
#             # matricula = Matricula.objects.filter(status=True, inscripcion=inscrito).exclude(nivel__periodo__nombre__icontains='TITULA').order_by('-id').first()
#             print(f'Inscrito: {postulante} - cedula: {inscrito.persona.cedula}')
#
#             inscripcion = Inscripcion(persona=postulante.inscripcionaspirante.persona,
#                                       fecha=datetime.now().date(),
#                                       carrera=consti,
#                                       modalidad=inscrito.modalidad,
#                                       sesion=inscrito.sesion,
#                                       sede=inscrito.sede,
#                                       colegio='',
#                                       aplica_b2=True,
#                                       fechainicioprimernivel=datetime.now().date(),
#                                       fechainiciocarrera=datetime.now().date(),
#                                       itinerario=1)
#             inscripcion.save()
#
#             print(f'Nueva inscripcion: {inscripcion} - Carrera: {inscripcion.carrera.id}')
#             postulante.inscripcion = inscripcion
#             postulante.save()
#             print(f'Postulación actualizada')
#             # matricula.inscripcion = inscripcion
#             # matricula.save()
#             # print(f'Matricula actualizada')
#
#             postulante.inscripcionaspirante.persona.crear_perfil(inscripcion=inscripcion)
#             a += 1
#             print(f'{a}/{inscritos.count()}')
#             if PerfilUsuario.objects.values_list('id').filter(inscripcion=inscripcion, status=True).exists():
#                 perfil = PerfilUsuario.objects.filter(inscripcion=inscripcion, status=True).last()
#                 perfil.inscripcionprincipal = True
#                 perfil.save()
#
#                 documentos = DocumentosDeInscripcion(inscripcion=inscripcion,
#                                                      titulo=False,
#                                                      acta=False,
#                                                      cedula=False,
#                                                      votacion=False,
#                                                      actaconv=False,
#                                                      partida_nac=False,
#                                                      pre=False,
#                                                      observaciones_pre='',
#                                                      fotos=False)
#                 documentos.save()
#                 preguntasinscripcion = inscripcion.preguntas_inscripcion()
#                 inscripciontesdrive = InscripcionTesDrive(inscripcion=inscripcion,
#                                                           licencia=False,
#                                                           record=False,
#                                                           certificado_tipo_sangre=False,
#                                                           prueba_psicosensometrica=False,
#                                                           certificado_estudios=False)
#                 inscripciontesdrive.save()
#                 inscripcion.malla_inscripcion()
#                 inscripcion.actualizar_nivel()
#                 if USA_TIPOS_INSCRIPCIONES:
#                     inscripciontipoinscripcion = InscripcionTipoInscripcion(inscripcion=inscripcion,
#                                                                             tipoinscripcion_id=TIPO_INSCRIPCION_INICIAL)
#                     inscripciontipoinscripcion.save()
#         else:
#             a += 1
# except Exception as ex:
#     pass
# try:
#     postulantes = InscripcionCohorte.objects.filter(id__in=[73001,
# 30368,
# 33074,
# 73749,
# 29765,
# 75710,
# 25808,
# 73525,
# 27970,
# 72495,
# 32770,
# 72680,
# 30232,
# 75933,
# 28700,
# 74399,
# 72587,
# 30563,
# 75707,
# 30002,
# 75923,
# 31498,
# 25872,
# 75383,
# 73673,
# 32348,
# 73705,
# 25691,
# 26412,
# 73977,
# 30546,
# 40701,
# 73676,
# 27109,
# 71379,
# 27613,
# 77657,
# 26483,
# 32768,
# 73808,
# 77027,
# 29366,
# 74238,
# 28110,
# 73634,
# 49529,
# 77428,
# 34580,
# 31489,
# 75922,
# 30414,
# 72813,
# 73793,
# 32134,
# 73843,
# 26903,
# 72951,
# 32126,
# 73512,
# 31215,
# 72545,
# 31944,
# 76076,
# 31953,
# 31952,
# 76069,
# 73330,
# 27007,
# 31654,
# 76486,
# 27190,
# 73964,
# 72567,
# 29687,
# 74054,
# 30325,
# 32130,
# 74061,
# 26106,
# 73802,
# 75051,
# 32272,
# 73884,
# 31988,
# 70563,
# 33723,
# 73798,
# 27831,
# 48911,
# 39435,
# 74311,
# 32431,
# 31936,
# 73913,
# 73134,
# 29778,
# 72772,
# 29962,
# 70560,
# 27981,
# 27803,
# 73801,
# 75187,
# 33722,
# 27864,
# 73754,
# 73963,
# 29515,
# 27424,
# 73603,
# 75974,
# 30500,
# 74048,
# 32091,
# 74252,
# 32921,
# 27211,
# 76535,
# 73767,
# 29608,
# 73882,
# 32124,
# 27173,
# 72547,
# 32097,
# 72392,
# 73677,
# 27676,
# 29734,
# 75771], inscripcion__isnull=False).order_by('inscripcionaspirante__persona__cedula')
#     c = 1
#     a = 1
#     lista1 = []
#     lista2 = []
#     for postulante in postulantes:
#         if Matricula.objects.filter(status=True, inscripcion=postulante.inscripcion).exclude(nivel__periodo__nombre__icontains='TITULA').exists():
#             if Matricula.objects.filter(inscripcion=postulante.inscripcion).exclude(nivel__periodo__nombre__icontains='TITULA').count() > 1:
#                 print(f'Consistentes: {c}/{postulantes.count()}')
#                 lista1.append(postulante.inscripcion.id)
#                 c += 1
#             else:
#                 print(f'Inconsistentes: {a}/{postulantes.count()}')
#                 lista2.append(postulante.inscripcion.id)
#                 a += 1
#     print(f'Matriculados2:{lista1}')
#     print(f'Matriculados1:{lista2}')
# except Exception as ex:
#     pass
# try:
#     #Excluir a los leads de la maestría vieja
#     idins = InscripcionCohorte.objects.filter(status=True, cohortes__id=194, inscripcion__isnull=False).values_list('inscripcion__id', flat=True)
#     inscripciones = Inscripcion.objects.filter(status=True, carrera__id=215, itinerario=2)
#     #nueva carrera
#     carrera = Carrera.objects.get(status=True, pk=266)
#     malla = Malla.objects.get(status=True, pk=500)
#     cont = 0
#     for inscripcion in inscripciones:
#         inscripcion.carrera = carrera
#         inscripcion.save()
#
#         for eInscripcionMalla in inscripcion.inscripcionmalla_set.all():
#             eInscripcionMalla.status = False
#             eInscripcionMalla.save()
#         eInscripcionMallas = inscripcion.inscripcionmalla_set.filter(inscripcion=inscripcion,
#                                                                      malla=malla)
#         if eInscripcionMallas.values("id").exists():
#             eInscripcionMalla = eInscripcionMallas[0]
#             eInscripcionMalla.status = True
#         else:
#             eInscripcionMalla = InscripcionMalla(inscripcion=inscripcion,
#                                                  malla=malla)
#         eInscripcionMalla.save()
#         inscripcion.actualizar_creditos()
#         inscripcion.actualizar_nivel()
#
#         inscor = None
#         if InscripcionCohorte.objects.filter(inscripcion=inscripcion).exists():
#             inscor = InscripcionCohorte.objects.filter(inscripcion=inscripcion).first()
#             if inscor.status == False:
#                 inscor.status = True
#             inscor.itinerario = 0
#             inscor.save()
#         cont += 1
#
#         print(f'{cont} - Cédula: {inscripcion.persona.cedula} - Inscrito: {inscripcion.persona} - Carrera: {inscripcion.carrera.id} - Itinerario: {inscor.itinerario} - {cont}/{inscripciones.count()}')
# except Exception as ex:
#     pass
# def imprimirreporte():
#     try:
#         estado = ''
#
#         __author__ = 'Unemi'
#
#         output = io.BytesIO()
#         name_document = 'reporte_rubros_inconsistentes'
#         nombre_archivo = name_document + "_6.xlsx"
#         directory = os.path.join(MEDIA_ROOT, 'reportes', 'gestion', nombre_archivo)
#         workbook = xlsxwriter.Workbook(directory, {'constant_memory': True})
#         ws = workbook.add_worksheet('resultados')
#         ws.set_column(0, 0, 10)
#         ws.set_column(1, 1, 15)
#         ws.set_column(2, 2, 15)
#         ws.set_column(3, 3, 30)
#         ws.set_column(4, 4, 30)
#         ws.set_column(5, 5, 40)
#         ws.set_column(6, 6, 30)
#         ws.set_column(7, 7, 30)
#
#         formatotitulo_filtros = workbook.add_format(
#             {'bold': 1, 'text_wrap': True, 'border': 1, 'align': 'center', 'font_size': 14})
#
#         formatoceldacab = workbook.add_format(
#             {'align': 'center', 'bold': 1, 'border': 1, 'text_wrap': True, 'fg_color': '#1C3247', 'font_color': 'white'})
#         formatoceldaleft = workbook.add_format(
#             {'text_wrap': True, 'align': 'center', 'valign': 'vcenter', 'border': 1})
#
#         formatoceldaleft2 = workbook.add_format(
#             {'num_format': '$ #,##0.00', 'text_wrap': True, 'align': 'center', 'valign': 'vcenter', 'border': 1, 'bold': 1})
#
#         formatoceldaleft3 = workbook.add_format(
#             {'text_wrap': True, 'align': 'right', 'valign': 'vcenter', 'border': 1, 'bold': 1})
#
#         decimalformat = workbook.add_format(
#             {'num_format': '$ #,##0.00', 'text_wrap': True, 'align': 'center', 'valign': 'vcenter', 'border': 1})
#
#         decimalformat2 = workbook.add_format(
#             {'num_format': '$ #,##0.00', 'text_wrap': True, 'align': 'center', 'valign': 'vcenter', 'border': 1, 'bold': 1})
#
#         ws.write(2, 0, 'N°', formatoceldacab)
#         ws.write(2, 1, 'Id', formatoceldacab)
#         ws.write(2, 2, 'Cédula', formatoceldacab)
#         ws.write(2, 3, 'Maestrante', formatoceldacab)
#         ws.write(2, 4, 'Rubro', formatoceldacab)
#         ws.write(2, 5, 'Fecha vencimiento', formatoceldacab)
#         ws.write(2, 6, 'Valor rubro', formatoceldacab)
#         ws.write(2, 7, 'Total pagado', formatoceldacab)
#         ws.write(2, 8, 'Cancelado', formatoceldacab)
#         ws.write(2, 9, 'Vencido', formatoceldacab)
#         ws.write(2, 10, '¿Activo?', formatoceldacab)
#
#
#         rubros = Rubro.objects.filter(status=True, cancelado=False, fecha__year__in=[2022, 2023, 2024], inscripcion__isnull=False)
#
#         filas_recorridas = 4
#         cont = 1
#
#         for rubro in rubros:
#             if rubro.total_pagado() > 0:
#                 ws.write('A%s' % filas_recorridas, str(cont), formatoceldaleft)
#                 ws.write('B%s' % filas_recorridas, str(rubro.id), formatoceldaleft)
#                 ws.write('C%s' % filas_recorridas, str(rubro.persona.cedula), formatoceldaleft)
#                 ws.write('D%s' % filas_recorridas, str(rubro.persona.nombre_completo_inverso()), formatoceldaleft)
#                 ws.write('E%s' % filas_recorridas, str(rubro.nombre), formatoceldaleft)
#                 ws.write('F%s' % filas_recorridas, str(rubro.fechavence), formatoceldaleft)
#                 ws.write('G%s' % filas_recorridas, rubro.valortotal, decimalformat)
#                 ws.write('H%s' % filas_recorridas, rubro.total_pagado(), decimalformat)
#                 ws.write('I%s' % filas_recorridas, str('Si' if rubro.cancelado else 'No'), formatoceldaleft)
#                 ws.write('J%s' % filas_recorridas, str('Si' if rubro.fechavence < datetime.now().date() else 'No'), formatoceldaleft)
#                 ws.write('K%s' % filas_recorridas, str('Si' if rubro.status else 'No'), formatoceldaleft)
#
#                 filas_recorridas += 1
#                 print(f'{cont} / {rubros.count()}')
#                 cont += 1
#             else:
#                 filas_recorridas += 1
#                 print(f'{cont} / {rubros.count()}')
#                 cont += 1
#
#         workbook.close()
#         # output.seek(0)
#         # fecha_hora_actual = datetime.now().date()
#         # filename = 'Listado_ventas_' + str(fecha_hora_actual) + '.xlsx'
#         response = HttpResponse(directory, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#         response['Content-Disposition'] = 'attachment; filename=%s' % name_document
#         url_file = "{}reportes/gestion/{}".format(MEDIA_URL, nombre_archivo)
#         print(url_file)
#     except Exception as ex:
#         pass
#
# imprimirreporte()

#         runjrcommand = [JR_JAVA_COMMAND, '-jar',
#                         os.path.join(JR_RUN, 'jasperstarter.jar'),
#                         'pr', eReporte.archivo.file.name,
#                         '--jdbc-dir', JR_RUN,
#                         '-f', tipo,
#                         '-t', 'postgres',
#                         '-H', DATABASES['sga_select']['HOST'],
#                         '-n', DATABASES['sga_select']['NAME'],
#                         '-u', DATABASES['sga_select']['USER'],
#                         '-p', f"'{DATABASES['sga_select']['PASSWORD']}'",
#                         '--db-port', DATABASES['sga_select']['PORT'],
#                         '-o', output_folder + os.sep + pdfname]
#         parametros = eReporte.parametros()
#         paramlist = [transform_jasperstarter(p, params) for p in parametros]
#         if paramlist:
#             runjrcommand.append('-P')
#             for parm in paramlist:
#                 if 'qr=' in parm and 'true' in parm:
#                     if eReporte.version == 2:
#                         url = pyqrcode.create(base_url + "/".join([MEDIA_URL, 'documentos', 'userreports', remove_accents(eUser.username), pdfname + "." + tipo]))
#                         url.png(os.path.join(os.path.join(SITE_STORAGE, 'media', 'qrcode', pdfname + '.png', 10, '#000000')))
#                     else:
#                         url = pyqrcode.create("http://sga.unemi.edu.ec" + "/".join([MEDIA_URL, 'documentos', 'userreports', remove_accents(eUser.username), pdfname + "." + tipo]))
#                         url.png('/var/lib/django/sistemagestion/media/qrcode/' + pdfname + '.png', 10, '#000000')
#                     runjrcommand.append(u'qr=' + pdfname + '.png')
#                 else:
#                     runjrcommand.append(parm)
#         else:
#             runjrcommand.append('-P')
#
#         runjrcommand.append(u'userweb=' + str(eUser.username))
#         if eReporte.version == 2:
#             runjrcommand.append(u'MEDIA_DIR=' + str("/".join([MEDIA_ROOT, ''])))
#             if DEBUG:
#                 runjrcommand.append(u'IMAGE_DIR=' + str("/".join([SUBREPOTRS_FOLDER, ''])))
#             else:
#                 runjrcommand.append(u'IMAGE_DIR=' + str(SUBREPOTRS_FOLDER))
#             runjrcommand.append(u'SUBREPORT_DIR=' + eReporte.ruta_subreport())
#             if isQR:
#                 runjrcommand.append(u'URL_QR=' + str(base_url + "/".join([MEDIA_URL, 'documentos', 'userreports', remove_accents(eUser.username), pdfname + "." + tipo])))
#                 runjrcommand.append(u'CODIGO_QR=' + str(codigo))
#                 runjrcommand.append(u'CERTIFICADO_ID=' + str(certificado.id))
#         else:
#             runjrcommand.append(u'SUBREPORT_DIR=' + str(SUBREPOTRS_FOLDER))
#         mens = ''
#         mensaje = ''
#         for m in runjrcommand:
#             mens += ' ' + m
#         urlbase = get_variable('SITE_URL_SGA')
#         reportfile = "/".join([MEDIA_URL, 'documentos', 'userreports', remove_accents(eUser.username), pdfname + "." + tipo])
#         # reportfile = f"{urlbase}{reportfile}"
#         if eReporte.es_background:
#             data['logreporte'] = logreporte
#             ReportBackground(params=params, data=data, reporte=eReporte, mensaje=mens, reportfile=reportfile, certificado=certificado, solicitud=eSolicitud).start()
#             lista_correos = []
#             if 'dirigidos' in params:
#                 dirigidos = json.loads(params['dirigidos'])
#                 personas = Persona.objects.filter(pk__in=dirigidos)
#                 for persona in personas:
#                     lista_correos.extend(persona.lista_emails())
#             if not 'no_persona_session' in params:
#                 if ePersona:
#                     lista_correos.extend(ePersona.lista_emails())
#             mensaje = 'El reporte se está realizando. Verifique los correos: %s después de unos minutos.' % (", ".join(lista_correos))
#         else:
#             if DEBUG:
#                 runjr = subprocess.run(mens, shell=True, check=True)
#                 # print('codigo:', mens)
#             else:
#                 runjr = subprocess.call(mens.encode("latin1"), shell=True)
#             if certificado:
#                 if certificado.funcionadjuntar:
#                     if certificado.funcionadjuntar.get_nombrefuncion_display() == 'adjuntar_malla_curricular':
#                         result, mensaje, reportfile_aux = certificado.funcionadjuntar.adjuntar_malla_curricular(logreporte)
#                         if not result:
#                             raise NameError(mensaje)
#                         reportfile = "/".join([MEDIA_URL, 'documentos', 'userreports', remove_accents(eUser.username), reportfile_aux])
#             reportfile = f"{urlbase}{reportfile}"
#             if eReporte.enviar_email:
#                 ids_personas = []
#                 if 'dirigidos' in params:
#                     dirigidos = json.loads(params['dirigidos'])
#                     personas = Persona.objects.filter(pk__in=dirigidos)
#                     for persona in personas:
#                         ids_personas.append(persona.id)
#                     if not 'no_persona_session' in params:
#                         if ePersona:
#                             ids_personas.append(ePersona.id)
#                 # reportfile = "/".join([MEDIA_URL, 'documentos', 'userreports', remove_accents(eUser.username), pdfname + "." + tipo])
#                 # reportfile = f"{urlbase}{reportfile}"
#                 for persona in Persona.objects.filter(pk__in=ids_personas).distinct():
#                     send_html_mail(eReporte.descripcion,
#                                    "reportes/emails/reporte_generacion.html",
#                                    {'sistema': u'SGA+',
#                                     'persona': persona,
#                                     'reporte': eReporte,
#                                     'reportfile': reportfile,
#                                     't': miinstitucion()},
#                                    persona.lista_emails(),
#                                    [],
#                                    cuenta=variable_valor('CUENTAS_CORREOS')[0]
#                                    )
#             if eSolicitud:
#                 if os.path.isfile(rutapdf):
#                     with open(rutapdf, 'rb') as f:
#                         archivo = f.read()
#                     buffer = io.BytesIO()
#                     buffer.write(archivo)
#                     pdf = buffer.getvalue()
#                     buffer.seek(0)
#                     buffer.close()
#                     fecha = datetime.now().date()
#                     hora = datetime.now().time()
#                     nombre_archivo_resultado = f"resultado_{fecha.year.__str__() + fecha.month.__str__() + fecha.day.__str__()}_{hora.hour.__str__() + hora.minute.__str__() + hora.second.__str__()}.pdf"
#                     eSolicitud.archivo_respuesta.save(nombre_archivo_resultado, ContentFile(pdf))
#                 eSolicitud.en_proceso = False
#
#                 if eSolicitud.perfil.inscripcion.carrera.coordinacion_carrera().id == 7 and eSolicitud.origen_object_id in [57, 39, 12]:
#                     eSolicitud.estado = 22
#                     eSolicitud.save()
#                 else:
#                     eSolicitud.estado = 2
#                     eSolicitud.save()
#                 observacion = f'Cambio de estado por proceso de entrega'
#                 #if ePersona.es_estudiante():
#                 #    observacion= f'Cambio de estado por generación'
#                 eHistorialSolicitud = HistorialSolicitud(solicitud=eSolicitud,
#                                                          observacion=observacion,
#                                                          fecha=datetime.now().date(),
#                                                          hora=datetime.now().time(),
#                                                          estado=eSolicitud.estado,
#                                                          responsable=ePersona,
#                                                          #archivo=eSolicitud.archivo_respuesta,
#                                                          )
#                 eHistorialSolicitud.save()
#                 titulo = f'Cambio de estado de la solicitud código {eSolicitud.codigo}'
#                 cuerpo = f'Solicitud código {eSolicitud.codigo} se encuentra en estado {eSolicitud.get_estado_display()}'
#                 eNotificacion = Notificacion(titulo=titulo,
#                                              cuerpo=cuerpo,
#                                              destinatario=eSolicitud.perfil.persona,
#                                              perfil=eSolicitud.perfil,
#                                              url='/alu_secretaria/mis_pedidos',
#                                              prioridad=1,
#                                              app_label='SIE',
#                                              fecha_hora_visible=datetime.now() + timedelta(days=2),
#                                              tipo=1,
#                                              en_proceso=False,
#                                              content_type=ContentType.objects.get_for_model(eSolicitud),
#                                              object_id=eSolicitud.pk,
#                                              )
#                 eNotificacion.save()
#
#         # sp = os.path.split(eReporte.archivo.file.name)
#         # return ok_json({'reportfile': "/".join([MEDIA_URL, 'documentos', 'userreports', remove_accents(eUser.username), pdfname + "." + tipo])})
#         aData = {'es_background': eReporte.es_background,
#                  'reportfile': reportfile
#                  }
#         return True, aData, mensaje
#     else:
#         tipo = params['rt']
#         d = datetime.now()
#         output_folder = os.path.join(JR_USEROUTPUT_FOLDER, remove_accents(eUser.username))
#         try:
#             os.makedirs(output_folder)
#         except Exception as ex:
#             pass
#         mensaje = ''
#
#         if tipo == 'pdf':
#             nombre = eReporte.nombre + d.strftime('%Y%m%d_%H%M%S') + ".pdf"
#             paramlist = {}
#             for p in eReporte.parametros():
#                 paramlist.update({p.nombre: fixparametro(p.tipo, params[p.nombre])})
#             global django_engine
#             django_engine = engines['django']
#             d = locals()
#             exec(eReporte.vista, globals(), d)
#             reportefinal = d['vistareporte'](eReporte, parametros=paramlist)
#             filepdf = open(output_folder + os.sep + nombre, "w+b")
#             pdf = pisa.pisaDocument(StringIO.BytesIO(reportefinal), dest=filepdf, link_callback=fetch_resources)
#             filepdf.close()
#         else:
#             nombre = eReporte.nombre + d.strftime('%Y%m%d_%H%M%S') + ".xls"
#             paramlist = {}
#             for p in eReporte.parametros():
#                 paramlist.update({p.nombre: fixparametro(p.tipo, params[p.nombre])})
#             d = locals()
#             exec(eReporte.vista, globals(), d)
#             book = d['vistareporte'](eReporte, parametros=paramlist)
#             output_folder = os.path.join(JR_USEROUTPUT_FOLDER, remove_accents(eUser.username))
#             filename = os.path.join(output_folder, nombre)
#             book.save(filename)
#         reportfile = "/".join([MEDIA_URL, 'documentos', 'userreports', remove_accents(eUser.username), nombre])
#         urlbase = get_variable('SITE_URL_SGA')
#         reportfile = f"{urlbase}{reportfile}"
#         aData = {'es_background': eReporte.es_background,
#                  'reportfile': reportfile
#                  }
#
# try:
#     eInscripciones  = Inscripcion.objects.filter(status=True, id__in=[58188])
#     for eInscripcion in eInscripciones:
#         ePerfilUsuario = PerfilUsuario.objects.get(status=True, inscripcion=eInscripcion)
#         eCoordinacion = eInscripcion.carrera.coordinacion_set.all()[0]
#         ePersona = ePerfilUsuario.persona
#         idencrypt = encrypt(eInscripcion.id)
#         parametros = {'n': 'rpt_certificado_egresado_grado', 'vqr': idencrypt, 'rt': 'pdf'}
#         eCertificado = Certificado.objects.get(reporte__id=667)
#         eServicio = eCertificado.servicio
#         valor = Decimal(eCertificado.costo).quantize(Decimal('.01'))
#         eContentTypeCertificado = ContentType.objects.get_for_model(eCertificado)
#         eSolicitudes = Solicitud.objects.filter(Q(en_proceso=True) | Q(estado__in=[1]), status=True,
#                                                 perfil_id=ePerfilUsuario.pk, servicio_id=eServicio.pk,
#                                                 origen_content_type_id=eContentTypeCertificado.pk,
#                                                 origen_object_id=eCertificado.id).exclude(servicio__proceso=8)
#         if not eSolicitudes.values("id").exists():
#             result = generar_codigo_solicitud(eServicio)
#             success = result.get('success', False)
#             codigo = result.get('codigo', None)
#             secuencia = result.get('secuencia', 0)
#             suffix = result.get('suffix', None)
#             prefix = result.get('prefix', None)
#             if not success:
#                 raise NameError(u"Código de solicitud no se pudo generar")
#             if codigo is None:
#                 raise NameError(u"Código de solicitud no se pudo generar")
#             if secuencia == 0:
#                 raise NameError(u"Secuenia del código de solicitud no se pudo generar")
#             if suffix is None:
#                 raise NameError(u"Sufijo del código de solicitud no se pudo generar")
#             if prefix is None:
#                 raise NameError(u"Prefijo del código de solicitud no se pudo generar")
#
#             eSolicitud = Solicitud(codigo=codigo,
#                                    secuencia=secuencia,
#                                    prefix=prefix,
#                                    suffix=suffix,
#                                    perfil_id=ePerfilUsuario.pk,
#                                    servicio_id=eServicio.pk,
#                                    origen_content_type=eContentTypeCertificado,
#                                    origen_object_id=eCertificado.pk,
#                                    descripcion=f'Solicitud ({codigo}) del Certificado académico con código {eCertificado.codigo} - {eCertificado.certificacion}',
#                                    fecha=datetime.now().date(),
#                                    hora=datetime.now().time(),
#                                    estado=22,
#                                    cantidad=1,
#                                    valor_unitario=valor,
#                                    subtotal=valor,
#                                    iva=0,
#                                    descuento=0,
#                                    en_proceso=False,
#                                    parametros=parametros,
#                                    tiempo_cobro=eCertificado.tiempo_cobro)
#             eSolicitud.save()
#             eHistorialSolicitud = HistorialSolicitud(solicitud=eSolicitud,
#                                                      observacion=eSolicitud.descripcion,
#                                                      fecha=eSolicitud.fecha,
#                                                      hora=eSolicitud.hora,
#                                                      estado=eSolicitud.estado,
#                                                      responsable=eSolicitud.perfil.persona,
#                                                      tipodocumento=3)
#             eHistorialSolicitud.save()
#
#             parametros = eSolicitud.parametros
#             ids_persona = [x.id for x in Persona.objects.filter(pk__in=[eSolicitud.perfil.persona_id])]
#             parametros['dirigidos'] = json.dumps(ids_persona)
#             parametros['no_persona_session'] = True
#             parametros['app'] = 'sie'
#
#             generar_certificado_digital(eSolicitud.perfil.persona, eCertificado.reporte, eSolicitud.perfil.persona.usuario, parametros, eSolicitud)
#             eSolicitud.estado=22
#             eSolicitud.save()
#             print(f'Certificado generado! - {eSolicitud.perfil.inscripcion.persona}')
# except Exception as ex:
#     print(ex)
#     print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))

    # fecha = datetime.now().date()
    # hora = datetime.now().time()
    # tipo = 'pdf'
    # paRequest = {
    #     'vqr': eInscripcion.id,
    # }
    # reporte = Reporte.objects.get(id=667)

    # d = run_report_v1(reporte=reporte, tipo=tipo, paRequest=paRequest)
    #
    # if not d['isSuccess']:
    #     raise NameError(d['mensaje'])
    # else:
    #     url_archivo = (SITE_STORAGE + d['data']['reportfile']).replace('\\', '/')
    #     url_archivo = (url_archivo).replace('//', '/')
    #     _name = f"resultado_{fecha.year.__str__() + fecha.month.__str__() + fecha.day.__str__()}_{hora.hour.__str__() + hora.minute.__str__() + hora.second.__str__()}.pdf"
    #     folder = os.path.join(SITE_STORAGE, 'media', 'archivo_respuesta', '')
    #     if not os.path.exists(folder):
    #         os.makedirs(folder)
    #     folder_save = os.path.join('archivo_respuesta', '').replace('\\', '/')
    #     url_file_generado = f'{folder_save}{_name}.pdf'
    #     ruta_creacion = SITE_STORAGE
    #     ruta_creacion = ruta_creacion.replace('\\', '/')
    #     shutil.copy(url_archivo, ruta_creacion + '/media/' + url_file_generado)
    #
    #     eSolicitud.archivo_respuesta = url_file_generado
    #     eSolicitud.estado = 22
    #     eSolicitud.save()
    #
    #     print(f'Certificado generado! - {eSolicitud.perfil.inscripcion.persona}')


    # tipootrorubro = 0
    # nuevorubro = TipoOtroRubro.objects.get(pk=3512)
    # rubros = Rubro.objects.filter(tipo_id=2925, fecha_creacion__year=2023).exclude(fecha_creacion__year__in=[2017, 2018, 2019, 2020, 2021, 2022])
    #
    # cursor = connections['epunemi'].cursor()
    # sql = """SELECT id FROM sagest_tipootrorubro WHERE idtipootrorubrounemi=%s; """ % (nuevorubro.id)
    # cursor.execute(sql)
    # registro = cursor.fetchone()
    #
    # # Si existe
    # if registro is not None:
    #     tipootrorubro = registro[0]
    # c = 0
    # for rubro in rubros:
    #     rubro.tipo = nuevorubro
    #     rubro.save(update_fields=['tipo'])
    #
    #     if rubro.idrubroepunemi != 0:
    #         # cursor = connections['epunemi'].cursor()
    #         sql = """SELECT id FROM sagest_rubro WHERE id=%s; """ % (rubro.idrubroepunemi)
    #         cursor.execute(sql)
    #         tienerubro = cursor.fetchone()
    #
    #         if tienerubro is not None:
    #             sql = """UPDATE sagest_rubro SET tipo_id=%s WHERE sagest_rubro.id=%s; """ % (tipootrorubro, rubro.idrubroepunemi)
    #             cursor.execute(sql)
    #     c += 1
    #     print(f'{c} - Rubro: {rubro.nombre} - id: {rubro.tipo.id} - idep: {tipootrorubro} - {rubro.tipo}')
    #
    # print(u"********************Nutrición***************************")
    #
    # tipootrorubro = 0
    # nuevorubro = TipoOtroRubro.objects.get(pk=3514)
    # rubros = Rubro.objects.filter(tipo_id=3233, fecha_creacion__year=2023)
    #
    # cursor = connections['epunemi'].cursor()
    # sql = """SELECT id FROM sagest_tipootrorubro WHERE idtipootrorubrounemi=%s; """ % (nuevorubro.id)
    # cursor.execute(sql)
    # registro = cursor.fetchone()
    #
    # # Si existe
    # if registro is not None:
    #     tipootrorubro = registro[0]
    # c = 0
    # for rubro in rubros:
    #     rubro.tipo = nuevorubro
    #     rubro.save(update_fields=['tipo'])
    #
    #     if rubro.idrubroepunemi != 0:
    #         # cursor = connections['epunemi'].cursor()
    #         sql = """SELECT id FROM sagest_rubro WHERE id=%s; """ % (rubro.idrubroepunemi)
    #         cursor.execute(sql)
    #         tienerubro = cursor.fetchone()
    #
    #         if tienerubro is not None:
    #             sql = """UPDATE sagest_rubro SET tipo_id=%s WHERE sagest_rubro.id=%s; """ % (tipootrorubro, rubro.idrubroepunemi)
    #             cursor.execute(sql)
    #     c += 1
    #     print(f'{c} - Rubro: {rubro.nombre} - id: {rubro.tipo.id} - idep: {tipootrorubro} - {rubro.tipo}')
    #
    # cursor.close()
# def imprimirreporte():
#     try:
#         estado = ''
#
#         __author__ = 'Unemi'
#
#         output = io.BytesIO()
#         name_document = 'reporte_mejores_puntuados'
#         nombre_archivo = name_document + "_4.xlsx"
#         directory = os.path.join(MEDIA_ROOT, 'reportes', 'gestion', nombre_archivo)
#         workbook = xlsxwriter.Workbook(directory, {'constant_memory': True})
#         ws = workbook.add_worksheet('resultados')
#         ws.set_column(0, 0, 5)
#         ws.set_column(1, 1, 35)
#         ws.set_column(2, 2, 30)
#         ws.set_column(3, 3, 35)
#         ws.set_column(4, 4, 15)
#         ws.set_column(5, 5, 30)
#         ws.set_column(6, 6, 30)
#         ws.set_column(7, 7, 20)
#         ws.set_column(8, 7, 15)
#         ws.set_column(9, 7, 20)
#         ws.set_column(10, 7, 20)
#         ws.set_column(11, 7, 15)
#         ws.set_column(12, 7, 15)
#         ws.set_column(13, 7, 15)
#         ws.set_column(14, 7, 15)
#         ws.set_column(15, 7, 15)
#
#         formatotitulo_filtros = workbook.add_format(
#             {'bold': 1, 'text_wrap': True, 'border': 1, 'align': 'center', 'font_size': 14})
#
#         formatoceldacab = workbook.add_format(
#             {'align': 'center', 'bold': 1, 'border': 1, 'text_wrap': True, 'fg_color': '#1C3247', 'font_color': 'white'})
#         formatoceldaleft = workbook.add_format(
#             {'text_wrap': True, 'align': 'center', 'valign': 'vcenter', 'border': 1})
#
#         formatoceldaleft2 = workbook.add_format(
#             {'num_format': '$ #,##0.00', 'text_wrap': True, 'align': 'center', 'valign': 'vcenter', 'border': 1, 'bold': 1})
#
#         formatoceldaleft3 = workbook.add_format(
#             {'text_wrap': True, 'align': 'right', 'valign': 'vcenter', 'border': 1, 'bold': 1})
#
#         decimalformat = workbook.add_format(
#             {'num_format': '$ #,##0.00', 'text_wrap': True, 'align': 'center', 'valign': 'vcenter', 'border': 1})
#
#         decimalformat2 = workbook.add_format(
#             {'num_format': '$ #,##0.00', 'text_wrap': True, 'align': 'center', 'valign': 'vcenter', 'border': 1, 'bold': 1})
#
#         ws.write(2, 0, 'N°', formatoceldacab)
#         ws.write(2, 1, 'Coordinacion', formatoceldacab)
#         ws.write(2, 2, 'Carrera', formatoceldacab)
#         ws.write(2, 3, 'Alumno', formatoceldacab)
#         ws.write(2, 4, 'Cédula', formatoceldacab)
#         ws.write(2, 5, 'Correo', formatoceldacab)
#         ws.write(2, 6, 'Correo institucional', formatoceldacab)
#         ws.write(2, 7, 'Teléfono', formatoceldacab)
#         ws.write(2, 8, 'Convencional', formatoceldacab)
#         ws.write(2, 9, 'Modalidad', formatoceldacab)
#         ws.write(2, 10, 'Nivel', formatoceldacab)
#         ws.write(2, 11, 'Materias malla', formatoceldacab)
#         ws.write(2, 12, 'Materias matricula', formatoceldacab)
#         ws.write(2, 13, 'Promedio', formatoceldacab)
#         ws.write(2, 14, 'Matriculado', formatoceldacab)
#         ws.write(2, 15, 'Paralelo', formatoceldacab)
#
#         matriculados = Matricula.objects.filter(status=True, nivelmalla__id=1, nivel__periodo__id=177, inscripcion__status=True,
#                                                 retiradomatricula=False, inscripcion__carrera__coordinacion__in=[1, 2, 3, 4, 5])
#
#         idcarreras = matriculados.values_list('inscripcion__carrera__id', flat=True).order_by('inscripcion__carrera__id').distinct()
#         carreras = Carrera.objects.filter(status=True, id__in=idcarreras).order_by('coordinacion__id')
#         filas_recorridas = 4
#         cont = 1
#
#         for carrera in carreras:
#             idins = matriculados.filter(inscripcion__carrera=carrera).values_list('inscripcion__id', flat=True).order_by('inscripcion__id').distinct()
#             idproms = RecordAcademico.objects.values_list('inscripcion__id', flat=True).filter(status=True,
#                                                                                              inscripcion__id__in=idins,
#                                                                                              modulomalla__isnull=True,
#                                                                                              asignaturamalla__isnull=False,
#                                                                                              noaplica=False).annotate(promedio_nota=Avg('nota')).order_by('-promedio_nota').exclude(inscripcion__id__in=[262976])
#             if idproms.count() > 0:
#                 c = 1
#                 for idprom in idproms:
#                     if c != 11:
#                         inscrito = Inscripcion.objects.get(pk=idprom)
#                         if inscrito.mi_nivel().nivel.id == 1:
#                             matriculado = Matricula.objects.filter(inscripcion__id=idprom, nivelmalla_id=1,
#                                                                    nivel__periodo_id=177, inscripcion__status=True,
#                                                                    retiradomatricula=False,
#                                                                    inscripcion__carrera__coordinacion__in=[1, 2, 3, 4, 5]).order_by('-id').first()
#
#                             asignaturas = AsignaturaMalla.objects.filter(status=True, malla=matriculado.inscripcion.mi_malla(), nivelmalla__id=1).count()
#                             idasignaturas = AsignaturaMalla.objects.filter(status=True, malla=matriculado.inscripcion.mi_malla(), nivelmalla__id=1).values_list('id', flat=True)
#                             matriculadas = MateriaAsignada.objects.filter(status=True, matricula=matriculado, materia__asignaturamalla__nivelmalla__id=1, materia__asignaturamalla__id__in=idasignaturas).count()
#
#                             curso = 'No registra'
#                             if MateriaAsignada.objects.filter(status=True, matricula=matriculado).exists():
#                                 curso = matriculado.materiaasignada_set.filter(matricula__status=True, status=True).first().materia.paralelo
#
#                             promedio = null_to_decimal(RecordAcademico.objects.filter(status=True,
#                                                                                       inscripcion=matriculado.inscripcion,
#                                                                                       asignaturamalla__isnull=False,
#                                                                                       noaplica=False).aggregate(promedio_nota=Avg('nota'))['promedio_nota'], 2)
#                             ws.write('A%s' % filas_recorridas, str(c), formatoceldaleft)
#                             ws.write('B%s' % filas_recorridas, str(matriculado.inscripcion.carrera.coordinacion_carrera().nombre), formatoceldaleft)
#                             ws.write('C%s' % filas_recorridas, str(matriculado.inscripcion.carrera.nombre), formatoceldaleft)
#                             ws.write('D%s' % filas_recorridas, str(matriculado.inscripcion.persona.nombre_completo_inverso()), formatoceldaleft)
#                             ws.write('E%s' % filas_recorridas, str(matriculado.inscripcion.persona.cedula), formatoceldaleft)
#                             ws.write('F%s' % filas_recorridas, str(matriculado.inscripcion.persona.email), formatoceldaleft)
#                             ws.write('G%s' % filas_recorridas, str(matriculado.inscripcion.persona.emailinst), formatoceldaleft)
#                             ws.write('H%s' % filas_recorridas, str(matriculado.inscripcion.persona.telefono), formatoceldaleft)
#                             ws.write('I%s' % filas_recorridas, str(matriculado.inscripcion.persona.telefono_conv), formatoceldaleft)
#                             ws.write('J%s' % filas_recorridas, str(matriculado.inscripcion.carrera.get_modalidad_display()), formatoceldaleft)
#                             ws.write('K%s' % filas_recorridas, str(matriculado.nivelmalla.nombre), formatoceldaleft)
#                             ws.write('L%s' % filas_recorridas, str(asignaturas), formatoceldaleft)
#                             ws.write('M%s' % filas_recorridas, str(matriculadas), formatoceldaleft)
#                             ws.write('N%s' % filas_recorridas, str(promedio), formatoceldaleft)
#                             ws.write('O%s' % filas_recorridas, str('SI'), formatoceldaleft)
#                             ws.write('P%s' % filas_recorridas, str(curso), formatoceldaleft)
#
#                             filas_recorridas += 1
#                             c += 1
#
#         workbook.close()
#         # output.seek(0)
#         # fecha_hora_actual = datetime.now().date()
#         # filename = 'Listado_ventas_' + str(fecha_hora_actual) + '.xlsx'
#         response = HttpResponse(directory, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#         response['Content-Disposition'] = 'attachment; filename=%s' % name_document
#         url_file = "{}reportes/gestion/{}".format(MEDIA_URL, nombre_archivo)
#         print(url_file)
#     except Exception as ex:
#         pass
#
# imprimirreporte()

# try:
#     idmaestrias = AsesorMeta.objects.filter(status=True, maestria__isnull=True, cohorte__procesoabierto=True).values_list('cohorte__maestriaadmision__id', flat=True).order_by('cohorte__maestriaadmision__id').distinct()
#     maestrias = MaestriasAdmision.objects.filter(status=True, id__in=idmaestrias)
#
#     asesores = AsesorComercial.objects.filter(status=True, activo=True)
#
#     for maestria in maestrias:
#         for asesor in asesores:
#             if not AsesorMeta.objects.filter(status=True, maestria=maestria, asesor=asesor).exists():
#                 if AsesorMeta.objects.filter(status=True, cohorte__maestriaadmision=maestria, asesor=asesor).exists():
#                     asesormeta = AsesorMeta(asesor=asesor,
#                                             maestria=maestria,
#                                             meta=0)
#                     asesormeta.save()
#                     print(f'{maestria} - Asesor asignado: {asesor}')
# except Exception as ex:
#     pass

    # inscripciones = InscripcionCohorte.objects.filter(status=True, vendido=False).order_by('id')
    # c = 0
    # for inscripcion in inscripciones:
    #     if Rubro.objects.filter(status=True, inscripcion=inscripcion, cancelado=True).exists():
    #         inscripcion.vendido = True
    #         inscripcion.save()
    #         if not VentasProgramaMaestria.objects.filter(status=True, inscripcioncohorte=inscripcion).exists():
    #             medio = ''
    #             if inscripcion.comprobante_subido():
    #                 medio = 'COMPROBANTE SUBIDO POR ASESOR'
    #                 fecha = inscripcion.get_comprobante_inscripcion().fecha_creacion.date()
    #                 hora = inscripcion.get_comprobante_inscripcion().fecha_creacion.time()
    #                 if inscripcion.asesor:
    #                     venta = VentasProgramaMaestria(inscripcioncohorte=inscripcion,
    #                                                    fecha=fecha,
    #                                                    hora=hora,
    #                                                    asesor=inscripcion.asesor,
    #                                                    mediopago=medio)
    #                     venta.save()
    #                 else:
    #                     venta = VentasProgramaMaestria(inscripcioncohorte=inscripcion,
    #                                                    fecha=fecha,
    #                                                    hora=hora,
    #                                                    mediopago=medio)
    #                     venta.save()
    #             elif inscripcion.comprobante_subido_epunemi():
    #                 medio = 'COMPROBANTE SUBIDO POR CONSULTA SALDOS'
    #                 fecha = inscripcion.get_comprobante_subido_epunemi().date()
    #                 hora = inscripcion.get_comprobante_subido_epunemi().time()
    #                 if inscripcion.asesor:
    #                     venta = VentasProgramaMaestria(inscripcioncohorte=inscripcion,
    #                                                    fecha=fecha,
    #                                                    hora=hora,
    #                                                    asesor=inscripcion.asesor,
    #                                                    mediopago=medio)
    #                     venta.save()
    #                 else:
    #                     venta = VentasProgramaMaestria(inscripcioncohorte=inscripcion,
    #                                                    fecha=fecha,
    #                                                    hora=hora,
    #                                                    mediopago=medio)
    #                     venta.save()
    #             elif inscripcion.tiene_pedidoonline_transferencia():
    #                 medio = 'PEDIDO ONLINE - TRANSFERENCIA'
    #                 fecha = inscripcion.fecha_pedidoonline_transferencia().date()
    #                 hora = inscripcion.fecha_pedidoonline_transferencia().time()
    #                 if inscripcion.asesor:
    #                     venta = VentasProgramaMaestria(inscripcioncohorte=inscripcion,
    #                                                    fecha=fecha,
    #                                                    hora=hora,
    #                                                    asesor=inscripcion.asesor,
    #                                                    mediopago=medio)
    #                     venta.save()
    #                 else:
    #                     venta = VentasProgramaMaestria(inscripcioncohorte=inscripcion,
    #                                                    fecha=fecha,
    #                                                    hora=hora,
    #                                                    mediopago=medio)
    #                     venta.save()
    #             elif inscripcion.tiene_pedidoonline_deposito():
    #                 medio = 'PEDIDO ONLINE - DEPOSITO'
    #                 fecha = inscripcion.fecha_pedidoonline_deposito().date()
    #                 hora = inscripcion.fecha_pedidoonline_deposito().time()
    #                 if inscripcion.asesor:
    #                     venta = VentasProgramaMaestria(inscripcioncohorte=inscripcion,
    #                                                    fecha=fecha,
    #                                                    hora=hora,
    #                                                    asesor=inscripcion.asesor,
    #                                                    mediopago=medio)
    #                     venta.save()
    #                 else:
    #                     venta = VentasProgramaMaestria(inscripcioncohorte=inscripcion,
    #                                                    fecha=fecha,
    #                                                    hora=hora,
    #                                                    mediopago=medio)
    #                     venta.save()
    #             elif inscripcion.tiene_pedidoonline_kushki():
    #                 medio = 'PEDIDO ONLINE - TARJETA DE CRÉDITO'
    #                 fecha = inscripcion.fecha_pedidoonline_kushki().date()
    #                 hora = inscripcion.fecha_pedidoonline_kushki().time()
    #                 if inscripcion.asesor:
    #                     venta = VentasProgramaMaestria(inscripcioncohorte=inscripcion,
    #                                                    fecha=fecha,
    #                                                    hora=hora,
    #                                                    asesor=inscripcion.asesor,
    #                                                    mediopago=medio)
    #                     venta.save()
    #                 else:
    #                     venta = VentasProgramaMaestria(inscripcioncohorte=inscripcion,
    #                                                    fecha=fecha,
    #                                                    hora=hora,
    #                                                    mediopago=medio)
    #                     venta.save()
    #             elif inscripcion.total_pagado_rubro_cohorte() > 0:
    #                 medio = 'VENTA DIRECTA DE CAJA'
    #                 fecha = inscripcion.fecha_primer_pago()
    #                 campodate = datetime.combine(fecha, time(12,0))
    #                 hora = campodate.time()
    #                 if inscripcion.asesor:
    #                     venta = VentasProgramaMaestria(inscripcioncohorte=inscripcion,
    #                                                    fecha=fecha,
    #                                                    hora=hora,
    #                                                    asesor=inscripcion.asesor,
    #                                                    mediopago=medio)
    #                     venta.save()
    #                     venta.facturado = True
    #                     venta.save()
    #                 else:
    #                     venta = VentasProgramaMaestria(inscripcioncohorte=inscripcion,
    #                                                    fecha=fecha,
    #                                                    hora=hora,
    #                                                    mediopago=medio)
    #                     venta.save()
    #                     venta.facturado = True
    #                     venta.save()
    #             c += 1
    #
    #             # if inscripcion.total_pagado_rubro_cohorte() > 0:
    #             #     ventaf = VentasProgramaMaestria.objects.filter(status=True, inscripcioncohorte=inscripcion).first()
    #             #     ventaf.facturado = True
    #             #     ventaf.save()
    #
    #             if inscripcion.asesor:
    #                 print(f'{c} - matriculado: {inscripcion.inscripcionaspirante.persona} - Cohorte: {inscripcion.cohortes.descripcion} - Maestria: {inscripcion.cohortes.maestriaadmision.descripcion} - Asesor: {inscripcion.asesor}')
    #             else:
    #                 print(f'{c} - matriculado: {inscripcion.inscripcionaspirante.persona} - Cohorte: {inscripcion.cohortes.descripcion} - Maestria: {inscripcion.cohortes.maestriaadmision.descripcion} - Asesor: NO REGISTRA')
# except Exception  as ex:
#     pass
# try:
#     matriculados = InscripcionCohorte.objects.filter(status=True, id__in=[20614, 20851, 21822, 22828, 22951, 23322, 23524, 23711, 23730, 24747, 24772,
#     24857, 24962, 25337, 25788, 25877, 25896, 26992, 27024, 28295, 28476, 28761, 29042, 32322, 32553, 32771, 35874, 35950, 37680, 38374, 38656, 39552,
#     39766, 39885, 39899, 40418, 41718, 44670, 49054, 50348, 51827, 53596, 36040, 36467, 38232, 39247, 40701, 41075, 41487, 48911])
#
#     for matriculado in matriculados:
#         rubros  = Rubro.objects.filter(status=True, inscripcion=matriculado, matricula__isnull=True)
#         matricula = Matricula.objects.get(status=True, inscripcion=matriculado.inscripcion)
#         if rubros:
#             for rubro in rubros:
#                 rubro.matricula = matricula
#                 rubro.save()
#             print(f"Rubros actualizados: {matriculado.inscripcionaspirante.persona} - Cohorte: {matriculado.cohortes}")
# except Exception as ex:
#     pass
# try:
#     matriculados = InscripcionCohorte.objects.filter(status=True, asesor__isnull=False)
#     c=1
#     for matriculado in matriculados:
#         if matriculado.cantidad_rubros_ins() > 0:
#             primerrubro = Rubro.objects.filter(status=True, inscripcion=matriculado).order_by('id').first()
#             if matriculado.formapagopac.id == 1:
#                 if matriculado.cohortes.valorprogramacertificado:
#                     if matriculado.cohortes.valorprogramacertificado == matriculado.total_pagado_rubro_cohorte():
#                         matriculado.vendido = True
#                         matriculado.save()
#                         print(f'N°-{c} - matriculado: {matriculado.inscripcionaspirante.persona} - Cohorte/Maestria: {matriculado.cohortes} - Asesor: {matriculado.asesor} - {matriculado.vendido}')
#                         c += 1
#                 else:
#                     if primerrubro.cancelado:
#                         matriculado.vendido = True
#                         matriculado.save()
#                         print(f'N°-{c} - matriculado: {matriculado.inscripcionaspirante.persona} - Cohorte/Maestria: {matriculado.cohortes} - Asesor: {matriculado.asesor} - {matriculado.vendido}')
#                         c += 1
#             else:
#                 if primerrubro.cancelado:
#                     matriculado.vendido = True
#                     matriculado.save()
#                     print(f'N°-{c} - matriculado: {matriculado.inscripcionaspirante.persona} - Cohorte/Maestria: {matriculado.cohortes} - Asesor: {matriculado.asesor} - {matriculado.vendido}')
#                     c += 1
# except Exception as ex:
#     print('error: %s' % ex)

#     cohorte2023 = CohorteMaestria.objects.get(id=151, status=True)
#     cont = 0
#     matriculados = []
#     fvence = datetime.strptime('2023-02-05', '%Y-%m-%d').date()
#
#     postu = InscripcionCohorte.objects.filter(status=True, cohortes__id=142)
#     for pos in postu:
#         if pos.tiene_matricula_cohorte():
#             matriculados.append(pos.id)
#
#     matriculados = InscripcionCohorte.objects.filter(status=True, estado_aprobador__in=[1, 2], cohortes__id=142).exclude(id__in=matriculados)
#     conta = 0
#     contado = 0
#     for matriculado in matriculados:
#         listarequisitos = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=matriculado, detalleevidenciarequisitosaspirante__estado_aprobacion=2, requisitos__requisito__claserequisito__clasificacion=1).values_list('requisitos__id', flat=True)
#         for lis in listarequisitos:
#             #COPIA A COLOR DE CERTIFICADO DE VOTACIÓN VIGENTE.
#             if lis in [896]:
#                 reqma = RequisitosMaestria.objects.get(status=True, pk=1048)
#                 evi = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=matriculado, requisitos__id=lis).order_by('-id').first()
#                 evi.requisitos = reqma
#                 evi.save()
#
#                 if EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=matriculado, requisitos__id=lis).count() > 1:
#                     evidences = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=matriculado, requisitos__id=lis).exclude(id=evi.id)
#                     for evid in evidences:
#                         evid.status = False
#                         evid.save()
#
#                 deta = DetalleEvidenciaRequisitosAspirante.objects.filter(status=True, evidencia=evi).order_by('-id')[0]
#                 if deta.fecha_aprobacion.date() < fvence:
#                     deta.observacion = 'EL CERTIFICADO DE VOTACIÓN NO ES EL VIGENTE'
#                     deta.observacion_aprobacion = 'EL CERTIFICADO DE VOTACIÓN NO ES EL VIGENTE'
#                     deta.estado_aprobacion = 3
#                     deta.save()
#             #CERTIFICADO DEL REGISTRO EN EL SENESCYT.
#             if lis in [894]:
#                 reqma = RequisitosMaestria.objects.get(status=True, pk=1046)
#                 evi = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=matriculado, requisitos__id=lis).order_by('-id').first()
#                 evi.requisitos = reqma
#                 evi.save()
#
#                 if EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=matriculado, requisitos__id=lis).count() > 1:
#                     evidences = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=matriculado, requisitos__id=lis).exclude(id=evi.id)
#                     for evid in evidences:
#                         evid.status = False
#                         evid.save()
#             #COPIA A COLOR DE CEDULA DE CIUDADANÍA O COPIA A COLOR DEL PASAPORTE EN CASO DE SER EXTRANJERO.
#             if lis in [895]:
#                 reqma = RequisitosMaestria.objects.get(status=True, pk=1047)
#                 evi = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=matriculado, requisitos__id=lis).order_by('-id').first()
#                 evi.requisitos = reqma
#                 evi.save()
#
#                 if EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=matriculado, requisitos__id=lis).count() > 1:
#                     evidences = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=matriculado, requisitos__id=lis).exclude(id=evi.id)
#                     for evid in evidences:
#                         evid.status = False
#                         evid.save()
#             #HOJA DE VIDA.
#             if lis in [897]:
#                 reqma = RequisitosMaestria.objects.get(status=True, pk=1049)
#                 evi = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=matriculado, requisitos__id=lis).order_by('-id').first()
#                 evi.requisitos = reqma
#                 evi.save()
#
#                 if EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=matriculado, requisitos__id=lis).count() > 1:
#                     evidences = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=matriculado, requisitos__id=lis).exclude(id=evi.id)
#                     for evid in evidences:
#                         evid.status = False
#                         evid.save()
#
#             #CERTIFICADO LABORAL CON EXPERIENCIA DOCENTE MÍNIMO 1 AÑO
#             if lis in [893]:
#                 reqma = RequisitosMaestria.objects.get(status=True, pk=1045)
#                 evi = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=matriculado, requisitos__id=lis).order_by('-id').first()
#                 evi.requisitos = reqma
#                 evi.save()
#
#                 if EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=matriculado, requisitos__id=lis).count() > 1:
#                     evidences = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=matriculado, requisitos__id=lis).exclude(id=evi.id)
#                     for evid in evidences:
#                         evid.status = False
#                         evid.save()
#
#         evidema = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=matriculado, detalleevidenciarequisitosaspirante__estado_aprobacion=2, requisitos__id__in=[1048, 1046, 1049, 1047, 1045])
#         if not evidema.count() == 5:
#             matriculado.estado_aprobador = 1
#             matriculado.save()
#             conta += 1
#         else:
#             if evidema.count() == 5 and matriculado.estado_aprobador == 1:
#                 matriculado.estado_aprobador = 2
#                 matriculado.save()
#             contado += 1
#
#         if matriculado.formapagopac and matriculado.formapagopac.id == 2:
#             listarequisitosfi = EvidenciaRequisitosAspirante.objects.filter(inscripcioncohorte=matriculado, requisitos__requisito__claserequisito__clasificacion=3, detalleevidenciarequisitosaspirante__estado_aprobacion=2)
#             for lis in listarequisitosfi:
#                 if RequisitosMaestria.objects.filter(requisito=lis.requisitos.requisito, cohorte=cohorte2023, status=True):
#                     requi = RequisitosMaestria.objects.filter(requisito=lis.requisitos.requisito,
#                                                               cohorte_id=cohorte2023,
#                                                               status=True)[0]
#                     lis.requisitos = requi
#                     lis.save()
#
#                     if lis.requisitos.requisito.id in [54, 62]:
#                         deta = DetalleEvidenciaRequisitosAspirante.objects.filter(status=True, evidencia=lis).order_by('-id').first()
#                         if deta.fecha_aprobacion.date() < fvence:
#                             deta.estado_aprobacion = 3
#                             deta.observacion = 'EL CERTIFICADO DE VOTACIÓN NO ES EL VIGENTE'
#                             deta.observacion_aprobacion = 'EL CERTIFICADO DE VOTACIÓN NO ES EL VIGENTE'
#                             deta.save()
#
#                         if DetalleEvidenciaRequisitosAspirante.objects.filter(status=True, evidencia=lis).count() > 1:
#                             detalles = DetalleEvidenciaRequisitosAspirante.objects.filter(status=True, evidencia=lis).exclude(id=deta.id)
#                             for det in detalles:
#                                 det.status = False
#                                 det.save()
#
#         if Rubro.objects.filter(inscripcion=matriculado, status=True).exists():
#             aspiranteconrubro = Rubro.objects.filter(inscripcion=matriculado, status=True)
#             for crubro in aspiranteconrubro:
#                 rubro = Rubro.objects.get(status=True, id=crubro.id)
#                 rubro.status = False
#                 rubro.save()
#
#                 if rubro.idrubroepunemi != 0:
#                     cursor = connections['epunemi'].cursor()
#                     sql = """SELECT id FROM sagest_pago WHERE rubro_id=%s; """ % (rubro.idrubroepunemi)
#                     cursor.execute(sql)
#                     tienerubropagos = cursor.fetchone()
#
#                     if tienerubropagos is None:
#                         sql = """DELETE FROM sagest_rubro WHERE sagest_rubro.id=%s AND sagest_rubro.idrubrounemi=%s; """ % (rubro.idrubroepunemi, rubro.id)
#                         cursor.execute(sql)
#                         cursor.close()
#
#         matriculado.cohortes_id = cohorte2023.id
#         # if matriculado.tiporespuesta:
#         #     matriculado.tiporespuesta = 1
#         matriculado.save()
#         cont += 1
#
#         estado = 0
#         if matriculado.estado_aprobador == 1:
#             estado = 'EN PROCESO'
#         elif matriculado.estado_aprobador == 2:
#             estado = 'ADMITIDO'
#         elif matriculado.estado_aprobador == 3:
#             estado = 'RECHAZADO'
#
#         print(cont, 'Lead:', ' ', matriculado.inscripcionaspirante.persona, 'Cedula:', ' ', matriculado.inscripcionaspirante.persona.cedula,'Cohorte:', ' ', matriculado.cohortes, 'Estado: ', estado, 'Canti: ', matriculado.total_evidencias())
# except Exception as ex:
#     print('error: %s' % ex)
#
# print("Finaliza")
#     aspirantes = InscripcionCohorte.objects.filter(id__in=[8843,
# 8841,
# 10075,
# 10125,
# 10954,
# 10115,
# 8652,
# 8598,
# 8460,
# 9374,
# 10113,
# 10419,
# 8475,
# 10158,
# 10892,
# 8706,
# 8688,
# 11070,
# 8746,
# 10272,
# 8464,
# 10958,
# 8447,
# 10567,
# 9375,
# 10425,
# 8580,
# 8614,
# 8608,
# 8726]).order_by(
#         'inscripcionaspirante__persona__apellido1', 'inscripcionaspirante__persona__apellido2',
#         'inscripcionaspirante__persona__nombres')
#     cont = 0
#     cohorte1 = CohorteMaestria.objects.get(id=98, status=True)
#     for inscripcioncoohorte in aspirantes:
#         listarequisitos = EvidenciaRequisitosAspirante.objects.filter(inscripcioncohorte=inscripcioncoohorte)
#         for lis in listarequisitos:
#             if RequisitosMaestria.objects.filter(requisito=lis.requisitos.requisito,
#                                                  cohorte_id=98, status=True):
#                 requi = RequisitosMaestria.objects.filter(requisito=lis.requisitos.requisito,
#                                                           cohorte_id=98,
#                                                           status=True)[0]
#                 lis.requisitos = requi
#                 lis.save()
#
#         if Rubro.objects.filter(inscripcion=inscripcioncoohorte, status=True).exists():
#             aspiranteconrubro = Rubro.objects.filter(inscripcion=inscripcioncoohorte, status=True)
#             for crubro in aspiranteconrubro:
#                 rubro = Rubro.objects.get(id=crubro.id)
#
#                 # if aspiranteconrubro.admisionposgradotipo == 2:
#                 #     tiporubroarancel = TipoOtroRubro.objects.get(pk=2845)
#                 #     rubro.nombre = tiporubroarancel.nombre + ' - ' + cohorte2.maestriaadmision.descripcion + ' - ' + cohorte2.descripcion
#                 # elif aspiranteconrubro.admisionposgradotipo == 3:
#                 #     rubro.nombre = cohorte2.maestriaadmision.descripcion + ' - ' + cohorte2.descripcion
#
#                 rubro.cohortemaestria = cohorte1
#                 rubro.save()
#
#         inscripcioncoohorte.cohortes_id = cohorte1.id
#         inscripcioncoohorte.save()
#         cont += 1
#         print(cont, 'Lead:', ' ', inscripcioncoohorte.inscripcionaspirante.persona, 'Cedula:', ' ', inscripcioncoohorte.inscripcionaspirante.persona.cedula,'Cohorte:', ' ', inscripcioncoohorte.cohortes)
#
#     aspirantes = InscripcionCohorte.objects.filter(id__in=[13886,
# 13141,
# 11063,
# 11951,
# 8907,
# 12889,
# 12416,
# 9144,
# 13006,
# 11320,
# 13610,
# 8795,
# 13504,
# 13115,
# 8456,
# 11686,
# 13047,
# 10632,
# 11066,
# 12863,
# 10651,
# 10418,
# 8927,
# 13808,
# 8492,
# 11615,
# 14070,
# 10447,
# 11755,
# 11558]).order_by(
#         'inscripcionaspirante__persona__apellido1', 'inscripcionaspirante__persona__apellido2',
#         'inscripcionaspirante__persona__nombres')
#     cont = 0
#     cohorte2 = CohorteMaestria.objects.get(id=102, status=True)
#     for inscripcioncoohorte in aspirantes:
#         listarequisitos = EvidenciaRequisitosAspirante.objects.filter(inscripcioncohorte=inscripcioncoohorte)
#         for lis in listarequisitos:
#             if RequisitosMaestria.objects.filter(requisito=lis.requisitos.requisito,
#                                                  cohorte_id=102, status=True):
#                 requi = RequisitosMaestria.objects.filter(requisito=lis.requisitos.requisito,
#                                                           cohorte_id=102,
#                                                           status=True)[0]
#                 lis.requisitos = requi
#                 lis.save()
#
#         if Rubro.objects.filter(inscripcion=inscripcioncoohorte, status=True).exists():
#             aspiranteconrubro = Rubro.objects.filter(inscripcion=inscripcioncoohorte, status=True)
#             for crubro in aspiranteconrubro:
#                 rubro = Rubro.objects.get(id=crubro.id)
#
#                 # if aspiranteconrubro.admisionposgradotipo == 2:
#                 #     tiporubroarancel = TipoOtroRubro.objects.get(pk=2845)
#                 #     rubro.nombre = tiporubroarancel.nombre + ' - ' + cohorte2.maestriaadmision.descripcion + ' - ' + cohorte2.descripcion
#                 # elif aspiranteconrubro.admisionposgradotipo == 3:
#                 #     rubro.nombre = cohorte2.maestriaadmision.descripcion + ' - ' + cohorte2.descripcion
#
#                 rubro.cohortemaestria = cohorte2
#                 rubro.save()
#
#         inscripcioncoohorte.cohortes_id = cohorte2.id
#         inscripcioncoohorte.save()
#         cont += 1
#         print(cont, 'Lead:', ' ', inscripcioncoohorte.inscripcionaspirante.persona, 'Cedula:', ' ', inscripcioncoohorte.inscripcionaspirante.persona.cedula,'Cohorte:', ' ', inscripcioncoohorte.cohortes)



# try:
#     miarchivo = openpyxl.load_workbook("contabilidad_financiamiento.xlsx")
#     lista = miarchivo.get_sheet_by_name('resultados')
#     totallista = lista.rows
#     a=0
#     c=0
#     for filas in totallista:
#         a += 1
#         if a > 1:
#             idinscripcioncohorte = str(filas[0].value)
#             # cedula = str((filas[6].value).replace(".", ""))
#             cedula = str((filas[6].value)).replace(".", "")
#             tipo = TipoFormaPagoPac.objects.get(id=2, status=True)
#             if idinscripcioncohorte != 'None':
#                 observacion = '30%+10 CUOTAS'
#                 if InscripcionCohorte.objects.filter(status=True, id=idinscripcioncohorte).exists():
#                     inscripcion = InscripcionCohorte.objects.get(status=True, id=idinscripcioncohorte)
#                     if AsesorComercial.objects.filter(status=True, persona__cedula=cedula).exists():
#                         asesor = AsesorComercial.objects.get(status=True, persona__cedula=cedula)
#
#                         inscripcion.formapagopac = tipo
#                         inscripcion.estadoformapago = 1
#                         inscripcion.save()
#
#                         deta = DetalleAprobacionFormaPago(inscripcion_id=inscripcion.id,
#                                                           formapagopac=tipo,
#                                                           estadoformapago=1,
#                                                           observacion=observacion,
#                                                           persona=asesor.persona)
#                         deta.save()
#
#                         c += 1
#                         print(c,' ','Lead:',inscripcion.inscripcionaspirante.persona,' - ','Asesor:',asesor, ' - ', 'FormaPago:', inscripcion.formapagopac.descripcion)
#                         # else:
#                         #     print(u"YA TIENE ASESOR:",'-', inscripcion.inscripcionaspirante.persona)
#                     else:
#                         print(u" NO EXISTE EL ASESOR", inscripcion.inscripcionaspirante.persona)
#                 else:
#                     print(u"-NO EXISTE EL LEAD")
#             else:
#                 print(u"FINALIZADO")
#                 break
# except Exception as ex:
#         print('error: %s' % ex)
#
# print("Finaliza")

# try:
#     # cohorte = CohorteMaestria.objects.get(id__in=[116, 118]).values_list('id')
#     id_inscripcion = Matricula.objects.filter(inscripcion__carrera__coordinacion__id=7).values_list('inscripcion__id', flat=True)
#
#     prospectos = InscripcionCohorte.objects.filter(status=True, cohortes__maestriaadmision__carrera__coordinacion__id=7,
#                            fecha_creacion__in=InscripcionCohorte.objects.values('inscripcionaspirante__id').annotate(fecha_creacion=Max('fecha_creacion')).values_list('fecha_creacion', flat=True).filter(status=True),
#                                                    estado_aprobador=2, fecha_creacion__lte = '2022-08-31').exclude(inscripcion__id__in=id_inscripcion)
#     tipo = TipoRespuestaProspecto.objects.get(id=5, status=True)
#     cont = 0
#     for prospecto in prospectos:
#         aspi = InscripcionCohorte.objects.get(pk=prospecto.id, status=True)
#
#         aspi.tiporespuesta = tipo
#         aspi.save()
#         cont += 1
#
#         print(cont, 'Lead:', ' ', aspi.inscripcionaspirante.persona, 'Respuesta:', ' ', aspi.tiporespuesta.descripcion)
#
# except Exception as ex:
#         print('error: %s' % ex)
#
# print("Finaliza")


# try:
#     # cohorte = CohorteMaestria.objects.get(id__in=[116, 118]).values_list('id')
#     id_inscripcion = Matricula.objects.filter(inscripcion__carrera__coordinacion__id=7).values_list('inscripcion__id', flat=True)
#
#     prospectos = InscripcionCohorte.objects.filter(status=True, cohortes__maestriaadmision__carrera__coordinacion__id=7,
#                            fecha_creacion__in=InscripcionCohorte.objects.values('inscripcionaspirante__id').annotate(fecha_creacion=Max('fecha_creacion')).values_list('fecha_creacion', flat=True).filter(status=True),
#                                                    estado_aprobador__in=[1, 3], fecha_creacion__lte = '2022-08-31').exclude(inscripcion__id__in=id_inscripcion, cohortes__id__in=[120, 123, 122, 113, 121, 86, 124, 107, 125, 126, 127, 128])
#     tipo = TipoRespuestaProspecto.objects.get(id=4, status=True)
#     cont = 0
#     for prospecto in prospectos:
#         aspi = InscripcionCohorte.objects.get(pk=prospecto.id, status=True)
#
#         aspi.tiporespuesta = tipo
#         aspi.save()
#         cont += 1
#
#         print(cont, 'Lead:', ' ', aspi.inscripcionaspirante.persona, 'Respuesta:', ' ', aspi.tiporespuesta.descripcion)
#
# except Exception as ex:
#         print('error: %s' % ex)
#
# print("Finaliza")

# try:
#     aspirantes = InscripcionCohorte.objects.filter(cohortes__id__in=[3, 32, 62, 74],
#                                                    status=True, estado_aprobador__in=[1,2]).order_by(
#         'inscripcionaspirante__persona__apellido1', 'inscripcionaspirante__persona__apellido2',
#         'inscripcionaspirante__persona__nombres')
#     cont = 0
#     for aspirante in aspirantes:
#         inscripcioncoohorte = InscripcionCohorte.objects.get(pk=aspirante.id)
#         listarequisitos = EvidenciaRequisitosAspirante.objects.filter(
#             inscripcioncohorte=inscripcioncoohorte)
#         for lis in listarequisitos:
#             if RequisitosMaestria.objects.filter(requisito=lis.requisitos.requisito,
#                                                  cohorte_id=86, status=True):
#                 requi = RequisitosMaestria.objects.filter(requisito=lis.requisitos.requisito,
#                                                           cohorte_id=86,
#                                                           status=True)[0]
#                 lis.requisitos = requi
#                 lis.save()
#
#         aspiranteconrubro = Rubro.objects.filter(inscripcion=inscripcioncoohorte, admisionposgradotipo__in=[2, 3], cohortemaestria_id=inscripcioncoohorte.cohortes_id, status=True)
#         for crubro in aspiranteconrubro:
#             crubro.cohortemaestria.id = 86
#             crubro.save()
#
#         inscripcioncoohorte.cohortes_id = 86
#         inscripcioncoohorte.save()
#         cont += 1
#         print(cont, 'Lead:', ' ', inscripcioncoohorte.inscripcionaspirante.persona, 'Cedula:', ' ', inscripcioncoohorte.inscripcionaspirante.persona.cedula,'Cohorte:', ' ', inscripcioncoohorte.cohortes)
#
# except Exception as ex:
#         print('error: %s' % ex)
#
# print("Finaliza")


# try:
#     miarchivo = openpyxl.load_workbook("contabilidad_financiamiento.xlsx")
#     lista = miarchivo.get_sheet_by_name('ventas')
#     totallista = lista.rows
#     a=0
#     for filas in totallista:
#         a += 1
#         if a > 1:
#             cedula = str(filas[9].value)
#             if cedula == 'None':
#                 cedula = ''
#             else:
#                 cedula = str((filas[9].value).replace(".", ""))
#             idinscripcioncohorte = str(filas[1].value)
#             if idinscripcioncohorte != 'None':
#                 observacion = 'MASIVO VENTAS COHORTE II BASICA'
#                 # urlzoom=u"https://unemi-edu-ec.zoom.us/j/%s"%idzoom
#                 # print(u"%s"%correo)
#                 if InscripcionCohorte.objects.filter(status=True, id=idinscripcioncohorte).exists():
#                     inscripcion = InscripcionCohorte.objects.get(status=True, id=idinscripcioncohorte)
#                     asesoranti = inscripcion.asesor
#                     if AsesorComercial.objects.filter(status=True,persona__cedula=cedula).exists():
#                         asesor = AsesorComercial.objects.get(status=True, persona__cedula=cedula)
#                         # if not inscripcion.asesor:
#                         inscripcion.asesor = asesor
#                         inscripcion.estado_asesor = 2
#                         inscripcion.save()
#
#                         if not asesoranti:
#                             histo = HistorialAsesor(inscripcion_id=inscripcion.id, fecha_inicio=inscripcion.fecha_modificacion,
#                                             fecha_fin=None, asesor=inscripcion.asesor, observacion=observacion)
#                             histo.save()
#                         else:
#                             if not asesoranti == asesor:
#                                 histoanti = HistorialAsesor.objects.get(inscripcion_id=inscripcion.id, fecha_fin=None)
#                                 histoanti.fecha_fin = inscripcion.fecha_modificacion
#                                 histoanti.save()
#                                 histo = HistorialAsesor(inscripcion_id=inscripcion.id, fecha_inicio=inscripcion.fecha_modificacion,
#                                                         fecha_fin=None, asesor=inscripcion.asesor, observacion=observacion)
#                                 histo.save()
#                             else:
#                                 print('El asesor ya tiene este lead asignado')
#
#                         print('Lead:',inscripcion.inscripcionaspirante.persona,' - ','Asesor:',asesor)
#                         # else:
#                         #     print(u"YA TIENE ASESOR:",'-', inscripcion.inscripcionaspirante.persona)
#                     else:
#                         print(u" NO EXISTE EL ASESOR", inscripcion.inscripcionaspirante.persona)
#                 else:
#                     print(u"-NO EXISTE EL LEAD")
#             else:
#                 print(u"FINALIZADO")
#                 break
# except Exception as ex:
#         print('error: %s' % ex)
#
# print("Finaliza")




