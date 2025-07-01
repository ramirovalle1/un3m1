# -*- coding: UTF-8 -*-
import json
import random
from datetime import datetime

from openpyxl import workbook as openxl
from openpyxl.styles import Font as openxlFont
from openpyxl.styles.alignment import Alignment as alin

from decimal import Decimal
from django.contrib.auth.decorators import login_required
from django.core import mail
from django.db import transaction
from django.db.models import Q
from django.http import HttpResponseRedirect, JsonResponse, HttpResponse
from django.shortcuts import render
from decorators import secure_module
from sagest.forms import MatrizRetencionForm
from sagest.models import ArchivoRetenciones, Proveedor
from sga.commonviews import adduserdata, obtener_reporte
from sga.funciones import MiPaginador, log, generar_nombre, variable_valor
from sga.models import miinstitucion, CUENTAS_CORREOS, Persona
from sga.tasks import send_html_mail, conectar_cuenta


@login_required(redirect_field_name='ret', login_url='/loginsagest')
@secure_module
@transaction.atomic()
def view(request):
    data = {}
    adduserdata(request, data)
    persona = request.session['persona']
    if request.method == 'POST':
        action = request.POST['action']

        if action == 'matrizlistaretencion':
            try:
                import openpyxl
                a = 0
                if 'archivomatriz' in request.FILES:
                    newfile = request.FILES['Hoja1']
                    if newfile.size > 2194304:
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"Error, Tamaño de archivo Maximo permitido es de 2Mb"})
                f = MatrizRetencionForm(request.POST)
                if f.is_valid():
                    miarchivo = openpyxl.load_workbook(request.FILES['Hoja1'])
                    hojas = miarchivo.get_sheet_names()
                    lista = miarchivo.get_sheet_by_name(str(hojas[0]))
                    totallista = lista.rows
                    for filas in totallista[:]:
                        a += 1
                        if a > 1:
                            if ArchivoRetenciones.objects.filter(numerocur=filas[0].value).exists():
                                retenciones = ArchivoRetenciones(numerocur=filas[0].value,
                                                                 proveedor=None)
                                retenciones.save(request)
                    return JsonResponse({'result': 'ok'})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', 'mensaje': u'Error al guardar los datos'})

        if action == 'addretencion':
            try:
                if 'archivopdf' in request.FILES:
                    newfile = request.FILES['archivopdf']
                    if newfile.size > 10485760:
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"Error, Tamaño de archivo Maximo permitido es de 10Mb"})
                    else:
                        newfiles = request.FILES['archivopdf']
                        newfilesd = newfiles._name
                        ext = newfilesd[newfilesd.rfind("."):]
                        if not ext.lower() == '.pdf':
                            return JsonResponse({"result": "bad", "mensaje": u"Error, solo archivos .pdf."})
                if 'archivoxml' in request.FILES:
                    newfilexml = request.FILES['archivoxml']
                    if newfilexml.size > 10485760:
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"Error, Tamaño de archivo Maximo permitido es de 10Mb"})
                    else:
                        newfilesxml = request.FILES['archivoxml']
                        newfilesdxml = newfilesxml._name
                        extxml = newfilesdxml[newfilesdxml.rfind("."):]
                        if not extxml.lower() == '.xml':
                            return JsonResponse({"result": "bad", "mensaje": u"Error, solo archivos .xml."})
                f = MatrizRetencionForm(request.POST, request.FILES)
                if f.is_valid():
                    retenciones = ArchivoRetenciones(proveedor=f.cleaned_data['proveedor'],
                                                     numerocur=f.cleaned_data['numerocur'],
                                                     comprobanteventa=f.cleaned_data['comprobanteventa'],
                                                     fechaemisionventa=f.cleaned_data['fechaemisionventa'],
                                                     comprobanteretencion=f.cleaned_data['comprobanteretencion'],
                                                     fechaemisionretencion=f.cleaned_data['fechaemisionretencion'],
                                                     montoretencion=f.cleaned_data['montoretencion'])
                    retenciones.save(request)
                    if 'archivopdf' in request.FILES:
                        newfile = request.FILES['archivopdf']
                        newfile._name = generar_nombre("retencionpdf_", newfile._name)
                        retenciones.archivopdf = newfile
                        retenciones.save(request)
                    if 'archivoxml' in request.FILES:
                        newfile = request.FILES['archivoxml']
                        newfile._name = generar_nombre("retencionxml_", newfile._name)
                        retenciones.archivoxml = newfile
                        retenciones.save(request)
                    log(u'Adiciono una nueva retencion: %s' % retenciones, request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as xe:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        if action == 'editretencion':
            try:
                if 'archivopdf' in request.FILES:
                    newfile = request.FILES['archivopdf']
                    if newfile.size > 10485760:
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"Error, Tamaño de archivo Maximo permitido es de 10Mb"})
                    else:
                        newfiles = request.FILES['archivopdf']
                        newfilesd = newfiles._name
                        ext = newfilesd[newfilesd.rfind("."):]
                        if not ext.lower() == '.pdf':
                            return JsonResponse({"result": "bad", "mensaje": u"Error, solo archivos .pdf."})
                if 'archivoxml' in request.FILES:
                    newfilexml = request.FILES['archivoxml']
                    if newfilexml.size > 10485760:
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"Error, Tamaño de archivo Maximo permitido es de 10Mb"})
                    else:
                        newfilesxml = request.FILES['archivoxml']
                        newfilesdxml = newfilesxml._name
                        extxml = newfilesdxml[newfilesdxml.rfind("."):]
                        if not extxml.lower() == '.xml':
                            return JsonResponse({"result": "bad", "mensaje": u"Error, solo archivos .xml."})
                f = MatrizRetencionForm(request.POST, request.FILES)
                if f.is_valid():
                    retencion = ArchivoRetenciones.objects.get(pk=request.POST['id'])
                    retencion.proveedor = f.cleaned_data['proveedor']
                    retencion.numerocur = f.cleaned_data['numerocur']
                    retencion.comprobanteventa = f.cleaned_data['comprobanteventa']
                    retencion.fechaemisionventa = f.cleaned_data['fechaemisionventa']
                    retencion.comprobanteretencion = f.cleaned_data['comprobanteretencion']
                    retencion.fechaemisionretencion = f.cleaned_data['fechaemisionretencion']
                    retencion.montoretencion = f.cleaned_data['montoretencion']
                    retencion.save(request)
                    if 'archivopdf' in request.FILES:
                        newfile = request.FILES['archivopdf']
                        newfile._name = generar_nombre("retencionpdf_", newfile._name)
                        retencion.archivopdf = newfile
                        retencion.save(request)
                    if 'archivoxml' in request.FILES:
                        newfile = request.FILES['archivoxml']
                        newfile._name = generar_nombre("retencionxml_", newfile._name)
                        retencion.archivoxml = newfile
                        retencion.save(request)
                    log(u'Edito una  retencion: %s' % retencion, request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos. %s" % ex})

        if action == 'delretencion':
            try:
                retencion = ArchivoRetenciones.objects.get(pk=request.POST['id'])
                retencion.status = False
                retencion.save()
                log(u'Elimino una retencion: %s' % retencion, request, "del")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        if action == 'envionotificacion':
            try:
                insretenciones = ArchivoRetenciones.objects.get(pk=request.POST['idinscripcioncohorte'])
                insretenciones.fecha_emailnotificacion = datetime.now()
                insretenciones.estado_emailnotificacion = True
                insretenciones.persona_envianotificacion = persona
                insretenciones.save(request)
                # per = Persona.objects.get(cedula='0923363030')
                log(u'Envio email envio retencion proveedor: %s' % (insretenciones), request, "add")
                if insretenciones.proveedor.email:
                    asunto = u"RETENCIONES - ARCHIVOS"
                    send_html_mail(asunto, "emails/notificar_retencion.html",
                                   {'sistema': request.session['nombresistema'], 'proveedor': insretenciones.proveedor,
                                    'insretenciones': insretenciones}, insretenciones.proveedor.emailpersonal(), [],
                                   [insretenciones.archivopdf, insretenciones.archivoxml, ],
                                   cuenta=CUENTAS_CORREOS[1][1])

                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al aprobar."})

        return HttpResponse(json.dumps({"result": "bad", "mensaje": u"Solicitud Incorrecta."}),
                            content_type="application/json")
    else:
        data['title'] = u'Retenciones'
        if 'action' in request.GET:
            action = request.GET['action']

            if action == 'matrizlistaretencion':
                try:
                    data['title'] = u'Subir listado de retenciones'
                    # form = MatrizListaRetencionForm()
                    # data['form'] = form
                    return render(request, "adm_archivo_retenciones/matrizlistaretencion.html", data)
                except Exception as ex:
                    pass

            if action == 'addretencion':
                try:
                    data['title'] = u'Adicionar Retención'
                    data['form'] = MatrizRetencionForm()
                    return render(request, 'adm_archivo_retenciones/addretencion.html', data)
                except Exception as ex:
                    pass

            if action == 'editretencion':
                try:
                    data['title'] = u'Editar Ubicación'
                    data['retencion'] = retencion = ArchivoRetenciones.objects.get(pk=request.GET['id'])
                    form = MatrizRetencionForm(initial={'proveedor': retencion.proveedor,
                                                        'numerocur': retencion.numerocur,
                                                        'comprobanteventa': retencion.comprobanteventa,
                                                        'fechaemisionventa': retencion.fechaemisionventa,
                                                        'comprobanteretencion': retencion.comprobanteretencion,
                                                        'fechaemisionretencion': retencion.fechaemisionretencion,
                                                        'montoretencion': retencion.montoretencion})
                    data['form'] = form
                    return render(request, "adm_archivo_retenciones/editretencion.html", data)
                except Exception as ex:
                    pass

            if action == 'delretencion':
                try:
                    data['title'] = u'Eliminar Retención'
                    data['retencion'] = ArchivoRetenciones.objects.get(pk=request.GET['id'])
                    return render(request, "adm_archivo_retenciones/delretencion.html", data)
                except Exception as ex:
                    pass

            return HttpResponseRedirect(request.path)
        else:
            search = None
            ids = None
            url_vars = ''
            filtro = Q(status=True)
            if 's' in request.GET:
                search = request.GET['s']
                url_vars += "&s={}".format(search)
            if search:
                if search.isdigit():
                    archivoretenciones = ArchivoRetenciones.objects.filter(Q(status=True) & (Q(numerocur=search) |
                                                                                             Q(comprobanteventa=search) |
                                                                                             Q(comprobanteretencion=search) |
                                                                                             Q(proveedor__nombre__icontains=search) |
                                                                                             Q(proveedor__alias__icontains=search) |
                                                                                             Q(proveedor__identificacion__icontains=search))).order_by('-numerocur')
                else:
                    archivoretenciones = ArchivoRetenciones.objects.filter(Q(status=True) &
                                                                           (Q(comprobanteventa=search) |
                                                                            Q(comprobanteretencion=search) |
                                                                            Q(proveedor__nombre__icontains=search) |
                                                                            Q(proveedor__alias__icontains=search) |
                                                                            Q(proveedor__identificacion__icontains=search))).order_by('-numerocur')
            elif 'id' in request.GET:
                ids = request.GET['id']
                archivoretenciones = ArchivoRetenciones.objects.filter(id=ids, status=True).order_by('-numerocur')
            else:
                archivoretenciones = ArchivoRetenciones.objects.filter(status=True).order_by('-fechaemisionretencion')
            paging = MiPaginador(archivoretenciones, 25)
            p = 1
            try:
                paginasesion = 1
                if 'paginador' in request.session:
                    paginasesion = int(request.session['paginador'])
                if 'page' in request.GET:
                    p = int(request.GET['page'])
                else:
                    p = paginasesion
                try:
                    page = paging.page(p)
                except:
                    p = 1
                page = paging.page(p)
            except:
                page = paging.page(p)
            request.session['paginador'] = p
            data['url_vars'] = url_vars
            data['paging'] = paging
            data['rangospaging'] = paging.rangos_paginado(p)
            data['page'] = page
            data['search'] = search if search else ""
            data['ids'] = ids if ids else ""
            data['archivoretenciones'] = page.object_list
            data['hoy'] = datetime.now().date()
            data['proveedores'] = Proveedor.objects.filter(status=True).distinct()
            data['reporte_3'] = obtener_reporte('resume_retenciones_tesoreria')
            data['reporte_4'] = obtener_reporte('resume_retenciones_tesoreria_proveedor')
            data['list_count'] = len(archivoretenciones)

            # EXPORTAR EXCEL PARA VISUALIZAR OFERTAS LABORALES
            if 'exportar_excel' in request.GET:
                if 'fechai' in request.GET and 'fechaf' in request.GET:
                    archivoretenciones = ArchivoRetenciones.objects.filter(status=True, fechaemisionretencion__gte=request.GET['fechai'], fechaemisionretencion__lte=request.GET['fechaf']).order_by('-fechaemisionretencion')
                elif 'proveedor' in request.GET:
                    archivoretenciones = ArchivoRetenciones.objects.filter(status=True, proveedor__id=request.GET['proveedor']).order_by('-fechaemisionretencion')
                    proveedorretencion = Proveedor.objects.values('identificacion','nombre').filter(status=True,id=request.GET['proveedor']).first()
                else:
                    archivoretenciones = ArchivoRetenciones.objects.filter(status=True).order_by('-fechaemisionretencion')
                wb = openxl.Workbook()
                wb["Sheet"].title = "Resumen_retenciones"
                ws = wb.active
                style_title = openxlFont(name='Arial', size=16, bold=True)
                style_cab = openxlFont(name='Arial', size=10, bold=True)
                alinear = alin(horizontal="center", vertical="center")
                response = HttpResponse(content_type="application/ms-excel")
                response[
                    'Content-Disposition'] = 'attachment; filename=Resumen de retenciones' + '-' + random.randint(
                    1, 10000).__str__() + '.xlsx'
                ws.column_dimensions['B'].width = 15
                ws.column_dimensions['C'].width = 15
                ws.column_dimensions['D'].width = 40
                ws.column_dimensions['E'].width = 30
                ws.column_dimensions['F'].width = 30
                ws.column_dimensions['G'].width = 30
                ws.column_dimensions['H'].width = 35
                ws.column_dimensions['I'].width = 25
                ws.column_dimensions['J'].width = 25

                ws.merge_cells('A1:J2')
                ws['A1'] = 'UNIVERSIDAD ESTATAL DE MILAGRO\n DEPARTAMENTO FINANCIERO SECCION TESORERIA'
                celda1 = ws['A1']
                celda1.font = style_title
                celda1.alignment = alinear

                ws.merge_cells('A3:B3')
                ws['A3'] = 'Resumen Retenciones'
                celda1 = ws['A3']
                celda1.font = style_cab
                if 'fechai' in request.GET and 'fechaf' in request.GET:
                    ws['A4'] = 'Desde'
                    celda1 = ws['A4']
                    celda1.font = style_cab

                    ws['B4'] = request.GET['fechai']
                    celda1 = ws['A4']
                    celda1.font = style_cab

                    ws['C4'] = 'Hasta'
                    celda1 = ws['C4']
                    celda1.font = style_cab

                    ws['D4'] = request.GET['fechaf']
                    celda1 = ws['C4']
                    celda1.font = style_cab

                elif 'proveedor' in request.GET:
                    ws['A4'] = 'RUC'
                    celda1 = ws['A4']
                    celda1.font = style_cab

                    ws['B4'] = proveedorretencion['identificacion']
                    celda1 = ws['A4']
                    celda1.font = style_cab

                    ws['C4'] = 'Proveedor'
                    celda1 = ws['C4']
                    celda1.font = style_cab

                    ws['D4'] = proveedorretencion['nombre']
                    celda1 = ws['C4']
                    celda1.font = style_cab
                # ws.merge_cells('A2:B3')
                # ws['A2'] = 'Empresa: ' + archivoretenciones.proveedor.identificacion
                # celda1 = ws['A2']
                # celda1.font = style_cab
                # celda1.alignment = alinear


                columns = [u"N°",u"N° CUR",u"RUC", u"NOMBRE", u"FEC. EMISION DE RETENCION", u"COMPROBANTE DE VENTA",
                           u"COMPROBANTE DE RETENCION", u"EMAILS",u"TELEFONOS",u"MONTO"
                           ]
                row_num = 5
                for col_num in range(0, len(columns)):
                    celda = ws.cell(row=row_num, column=(col_num + 1), value=columns[col_num])
                    celda.font = style_cab
                row_num = 6
                numero = 1
                for list in archivoretenciones:
                    ws.cell(row=row_num, column=1, value=numero)
                    ws.cell(row=row_num, column=2, value=str(list.numerocur) if list.numerocur else '0')
                    ws.cell(row=row_num, column=3, value=str(list.proveedor.identificacion))
                    ws.cell(row=row_num, column=4, value=str(list.proveedor.nombre))
                    ws.cell(row=row_num, column=5,
                             value=str(list.fechaemisionretencion) if list.fechaemisionretencion else 'Sin asignar')
                    ws.cell(row=row_num, column=6, value=str(list.comprobanteventa)if list.comprobanteventa else '')
                    ws.cell(row=row_num, column=7, value=str(list.comprobanteretencion)if list.comprobanteretencion else '')
                    ws.cell(row=row_num, column=8,
                            value=str(list.proveedor.email))
                    telf = []
                    for telefono in list.proveedor.lista_telefonos():
                        telf.append(telefono)
                    ws.cell(row=row_num, column=9,
                            value=", ".join(telf))
                    ws.cell(row=row_num, column=10,
                            value=str(list.montoretencion)if list.montoretencion else '0')


                    # ws.cell(row=row_num, column=4,
                    #         value=str(NIVEL_INSTRUCCION[0][1]) if list.nivel == 3 else str(NIVEL_INSTRUCCION[1][1]))
                    # ws.cell(row=row_num, column=5, value=str(MODALIDAD[list.modalidad][1]))
                    # ws.cell(row=row_num, column=6, value=str(DEDICACION[list.dedicacion][1]))
                    # ws.cell(row=row_num, column=7, value=str(JORNADA[list.jornada][1]))
                    # ws.cell(row=row_num, column=8, value=str(list.rmu))
                    # ws.cell(row=row_num, column=9, value=str(list.tipocontrato.nombre))
                    # ws.cell(row=row_num, column=10, value=str(TIEMPO[list.tiempoexperiencia][1]))
                    # ws.cell(row=row_num, column=11, value=str(list.vacantes))
                    # ws.cell(row=row_num, column=12, value=str(list.finicio) if list.finicio else 'Sin asignar')
                    # ws.cell(row=row_num, column=13,
                    #         value=str(list.ffin) if list.ffin else 'Sin asignar')
                    # ws.cell(row=row_num, column=14,
                    #         value=str(list.finiciopostulacion) if list.finiciopostulacion else 'Sin asignar')
                    # ws.cell(row=row_num, column=15,
                    #         value=str(list.ffinpostlacion) if list.ffinpostlacion else 'Sin asignar')
                    # ws.cell(row=row_num, column=16,
                    #         value=str(list.finiciorevision) if list.finiciorevision else 'Sin asignar')
                    # ws.cell(row=row_num, column=17,
                    #         value=str(list.ffinrevision) if list.ffinrevision else 'Sin asignar')
                    # ws.cell(row=row_num, column=5,
                    #         value=str(list.persona.emailinst) if list.persona.emailinst else list.persona.email)
                    # ws.cell(row=row_num, column=6, value=str(list.fecha_creacion.date()))
                    # ws.cell(row=row_num, column=7,
                    #         value=str(list.fecha_revision) if list.fecha_revision else 'Sin revisar')
                    # ws.cell(row=row_num, column=8, value=str(list.get_estado_display()))
                    row_num += 1
                    numero += 1
                wb.save(response)
                return response

            return render(request, "adm_archivo_retenciones/view.html", data)
