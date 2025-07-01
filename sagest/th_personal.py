# -*- coding: UTF-8 -*-
import os
import random
import sys
import io
import zipfile
import json
import xlsxwriter
import xlwt
from openpyxl import workbook as openxl
from openpyxl.chart import ScatterChart, Reference, Series, PieChart, BarChart
from openpyxl.styles import Font as openxlFont
from openpyxl.styles.alignment import Alignment as alin
from django.contrib import messages
from django.db.models import Avg, Exists, OuterRef, Subquery, Count
from openpyxl import load_workbook
from unidecode import unidecode

from med.models import PersonaExtension, PersonaExamenFisico
from pdip.models import ActividadesPerfil, ContratoDip, ContratoCarrera
from sga.excelbackground import descarga_masica_documentos_familiares_background, descarga_cargas_gastosp_background, \
    descarga_masiva_actas_firmadas_background, descarga_masiva_requisitos_pazsalvo_background, descarga_masiva_requisitos_ingreso_background, \
descargar_comprimidos_fotos_perfi_jpg_background
from sga.funcionesxhtml2pdf import conviert_html_to_pdfsave, conviert_html_to_pdf, conviert_html_to_pdf_name
import pyqrcode
import code128
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Q
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.shortcuts import render, redirect
from django.template import Context
from django.template.loader import get_template
from django.core.files import File
from xlwt import *
from django.forms import model_to_dict
from decorators import secure_module, last_access
from sagest.forms import DatosPersonalesForm, DatosNacimientoForm, DatosDomicilioForm, DatosMedicosForm, \
    ContactoEmergenciaForm, EtniaForm, DiscapacidadForm, FamiliarForm, DeclaracionBienForm, CuentaBancariaPersonaForm, \
    TitulacionPersonaForm, ArchivoTitulacionForm, CapacitacionPersonaForm, ArchivoCapacitacionForm, \
    ExperienciaLaboralForm, DatosPersonaInstitucionForm, ArchivoIdiomaForm, ArchivoExperienciaForm, \
    PersonaContratosForm, PersonaAccionesForm, CertificadoPersonaForm, PersonaDetalleMaternidad, \
    ArchivoHistorialContratoDirectorForm, ArchivoHistorialContratoForm, AccionPersonalDocumentoForm, \
    PersonaAccionesOldForm, BitacoraForm, DescargarCompromidoForm, DescagarCargasFechasForm, SubirPermisoMasivoForm, \
    ImportarArchivoForm, SagestImportarXLSForm, PeriodoTTHHForm, RequisitoPeriodoTTHHForm, ValidarRequisitoTTHHForm, \
    ComprimidoIngresoForm, MigranteForm, RedAcademicaForm, ParRevisorArticuloForm, DetalleTitulacionBachillerForm, \
    DescargarContratosForm, DeclaracionesPersonalForm
from sagest.models import ExperienciaLaboral, RolPago, PersonaContratos, PersonaAcciones, SolicitudPublicacion, \
    ActivoFijo, OtroMerito, DistributivoPersona, RegimenLaboral, DistributivoPersonaHistorial, \
    HistorialArchivosContratos, HistoricoDocumentosPersonaAcciones, AccionPersonal, Departamento, TIPO_SISTEMA, \
    TIPO_ACTIVIDAD_BITACORA, GastosPersonales, PeriodoGastosPersonales, RegistroDecimo, ConfiguraDecimo, \
    ConfiguraPeriodotthh, Ubicacion, \
    RequisitoPeriodotthh, PersonaPeriodotthh, DocumentoPersonaPeriodotthh, ESTADOS_DOCUMENTOS_REQUISITOS, \
    ResumenMesGastosPersonales, CapEventoPeriodo, CapCabeceraSolicitud, VacunaCovid
from settings import PROFESORES_GROUP_ID, EMPLEADORES_GRUPO_ID, ALUMNOS_GROUP_ID, SITE_STORAGE, PUESTO_ACTIVO_ID, MEDIA_ROOT,DEBUG, MEDIA_URL
from sga.commonviews import adduserdata, obtener_reporte
from sga.forms import IdiomaDominaForm, PeriodoSabaticoForm, EvidenciaForm, PersonaAportacionHistorialLaboralForm, \
    ReferenciaPersonaForm, CertificadoIdiomaForm
from sga.funciones import MiPaginador, log, generar_nombre, variable_valor, puede_realizar_accion, convertir_fecha, \
    convertir_fecha_invertida, null_to_decimal, remover_caracteres_especiales_unicode, convertir_fecha_hora_invertida, \
    puede_realizar_accion_afirmativo, puede_realizar_accion_is_superuser,notificacion
from sga.models import Persona, PersonaAportacionHistorialLaboral, PersonaDatosFamiliares, DeclaracionBienes, \
    CuentaBancariaPersona, Titulacion, \
    Capacitacion, IdiomaDomina, Persona, ArticuloInvestigacion, PonenciasInvestigacion, LibroInvestigacion, Evidencia, \
    NivelTitulacion, ParticipantesMatrices, RespuestaEvaluacionAcreditacion, ResumenFinalProcesoEvaluacionIntegral, \
    MigracionEvaluacionDocente, ResumenParcialEvaluacionIntegral, null_to_numeric, ResumenFinalEvaluacionAcreditacion, \
    CapituloLibroInvestigacion, ResponsableEvaluacion, CertificacionPersona, ProcesoEvaluativoAcreditacion, \
    ParticipantesArticulos, ArticulosBaseIndexada, CertificadoTutoriaHV, CoordinadorCarrera, Profesor, ProfesorMateria, \
    ClaseActividad, MESES_CHOICES, Materia, Notificacion, MigrantePersona, PersonaDocumentoPersonal, PerfilUsuario, \
    Externo, CUENTAS_CORREOS, ReferenciaPersona, Titulo, Archivo, RedPersona, ParRevisorProduccionCientifica, \
    DetalleTitulacionBachiller, CapCabeceraSolicitudDocente, Graduado, CertificadoIdioma, \
    TribunalTemaTitulacionPosgradoMatricula
from datetime import datetime, timedelta
import calendar
from PIL import Image
import requests

from sagest.models import BitacoraActividadDiaria
from sagest.funciones import dominio_sistema_base, encrypt_id, ext_archive, departamentos_vigentes, get_departamento
from sga.tasks import send_html_mail
from sga.templatetags.sga_extras import encrypt
from utils.filtros_genericos import filtro_persona

from core.choices.models.sagest import ESTADO_INCIDENCIA, ETAPA_INCIDENCIA, ESTADO_SANCION_PERSONA, ROL_FIRMA_DOCUMENTO
from core.firmar_documentos import obtener_posicion_x_y_saltolinea
from core.firmar_documentos_ec import JavaFirmaEc

import openpyxl

@login_required(redirect_field_name='ret', login_url='/loginsagest')
@secure_module
@last_access
@transaction.atomic()
def view(request):
    data = {}
    adduserdata(request, data)
    persona = request.session['persona']
    administrativo = None
    usuario = request.user
    dominio_sistema=dominio_sistema_base(request)
    if request.method == 'POST':
        action = request.POST['action']

        if action == 'pdffichacatalograficas_articulo':
            try:
                data = {}
                data['participantess'] = articulo = ParticipantesArticulos.objects.filter(articulo__id=request.POST['id'], status=True)
                data['cantidad'] = articulo.count()
                data['basesindexadas'] = ArticulosBaseIndexada.objects.filter(articulo_id=request.POST['id'],status=True)
                return conviert_html_to_pdf(
                    'inv_articulos/fichacatalografica_articulo_pdf.html',
                    {
                        'pagesize': 'A4',
                        'data': data,
                    }
                )
            except Exception as ex:
                pass

        elif action == 'detallecertificaciontutoria':
            try:
                data['certificacion'] = certificacion = CertificadoTutoriaHV.objects.get(pk=int(request.POST['id']))
                if certificacion.usuario_creacion:
                    data['personacreacion'] = Persona.objects.get(
                        usuario=certificacion.usuario_creacion) if certificacion.usuario_creacion.id > 1 else ""
                template = get_template("th_hojavida/detallecertificaciontutoria.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'html': json_content})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'detalledato':
            try:
                data = {}
                personaladministrativo = Persona.objects.filter(Q(administrativo__isnull=False) | Q(profesor__isnull=False)).distinct()
                data['listaadministrativos'] = personaladministrativo.filter(Q(titulacion__verificado=False, titulacion__cursando=False, titulacion__archivo__isnull=False, titulacion__registroarchivo__isnull=False) |
                                                                             Q(capacitacion__verificado=False, capacitacion__archivo__isnull=False) |
                                                                             Q(declaracionbienes__verificado=False, declaracionbienes__archivo__isnull=False) |
                                                                             Q(cuentabancariapersona__verificado=False, cuentabancariapersona__archivo__isnull=False) |
                                                                             Q(experiencialaboral__verificado=False, experiencialaboral__archivo__isnull=False)).distinct()
                template = get_template("th_personal/detalle.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'data': json_content})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

    #Información laboral

        # Aportación laboral

        elif action == 'addaportacionhistoriallaboral':
            try:
                id = int(request.POST['id'])
                formvalor = (encrypt(request.POST['idp']).replace('0', ''))
                f = PersonaAportacionHistorialLaboralForm(request.POST, request.FILES)
                newfile = None
                if 'archivo_resumen' in request.FILES:
                    newfile = request.FILES['archivo_resumen']
                    if newfile:
                        newfilesd = newfile._name
                        ext = newfilesd[newfilesd.rfind("."):]
                        if not ext == '.pdf':
                            return JsonResponse({"result": "bad", "mensaje": u"Error, solo archivos .pdf."})
                        if newfile.size > 20480000:
                            return JsonResponse({"result": "bad", "mensaje": u"Error, archivo mayor a 20 Mb."})
                        if newfile:
                            newfile._name = generar_nombre("archivo_resumen", newfile._name)
                newfile2 = None
                if 'archivo_detalle' in request.FILES:
                    newfile2 = request.FILES['archivo_detalle']
                    if newfile2:
                        newfilesd2 = newfile2._name
                        ext2 = newfilesd2[newfilesd2.rfind("."):]
                        if not ext2 == '.pdf':
                            return JsonResponse({"result": "bad", "mensaje": u"Error, solo archivos .pdf."})
                        if newfile2.size > 20480000:
                            return JsonResponse({"result": "bad", "mensaje": u"Error, archivo mayor a 20 Mb."})
                        if newfile2:
                            newfile2._name = generar_nombre("archivo_detalle", newfile2._name)
                if not f.is_valid():
                    for k, v in f.errors.items():
                        raise NameError(v[0])
                if formvalor == 'add':
                    registro = PersonaAportacionHistorialLaboral(persona_id=id)
                elif formvalor == 'edit':
                    registro = PersonaAportacionHistorialLaboral.objects.get(pk=id)
                if newfile:
                    registro.archivo_resumen = newfile
                if newfile2:
                    registro.archivo_detalle = newfile2
                registro.save(request)
                log(u'%s nuevo Historial Laboral de Aportación: %s' % (
                    'Adiciono' if formvalor == 'add' else 'Edito', registro), request, formvalor)
                return JsonResponse({"result": False, "mensaje": u"Registro guardado exitosamente."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse(
                    {"result": True, "mensaje": u"Error al guardar los datos. %s" % ex.__str__()})

        elif action == 'delaportacionhistoriallaboral':
            try:
                accion = PersonaAportacionHistorialLaboral.objects.get(pk=encrypt(request.POST['id']))
                accion.status = False
                accion.save(request)
                res_json = {"error": False}
            except Exception as ex:
                transaction.set_rollback(True)
                res_json = {'error': True, "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)

                # Situación laboral

                # Rol de pago

        elif action == 'reportedeclaraciones':
            try:
                __author__ = 'Unemi'
                style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
                style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
                style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
                title = easyxf(
                    'font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                style1 = easyxf(num_format_str='D-MMM-YY')
                font_style = XFStyle()
                font_style.font.bold = True
                font_style2 = XFStyle()
                font_style2.font.bold = False
                font_style_center = XFStyle()
                font_style_center.alignment.horz = Alignment.HORZ_CENTER
                font_style_center.font.bold = False
                wb = Workbook(encoding='utf-8')
                ws = wb.add_sheet('reporte')
                ws.write_merge(0, 0, 0, 10, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                columns = [
                    (u"FECHA CREACIÓN", 4000),
                    (u"CEDULA", 4000),
                    (u"EMPLEADO", 9000),
                    (u"TIPO DECLARACION", 4000),
                    (u"FECHA DECLARACION", 4000),
                    (u"FECHA INICIO PERIODO", 4000),
                    (u"CÓDIGO BARRA", 4000),
                    (u"DEPARTAMENTO", 9000),
                    (u"CARGO", 9000),
                    (u"ARCHIVO", 9000),
                    (u"ESTADO ACTUAL PERSONA", 4000),
                ]
                row_num = 1
                for col_num in range(len(columns)):
                    ws.write(row_num, col_num, columns[col_num][0], font_style)
                    ws.col(col_num).width = columns[col_num][1]
                date_format = xlwt.XFStyle()
                date_format.num_format_str = 'yyyy-mm-dd'
                date_formatreverse = xlwt.XFStyle()
                date_formatreverse.num_format_str = 'dd/mm/yyyy'

                filtro = Q(status=True)
                fdesde, fhasta = request.POST.get('fachadesde', ''), request.POST.get( 'fachahasta', '')
                tipodeclaracion, estado = request.POST.get('tipodeclaracion', '0'), request.POST.get('estado', '0')
                if fdesde:
                    filtro = filtro & Q(fecha__gte=fdesde)
                if fhasta:
                    filtro = filtro & Q(fecha__lte=fhasta)
                if tipodeclaracion != '0':
                    filtro = filtro & Q(tipodeclaracion=tipodeclaracion)

                if estado != '0':
                    ids_personas = DeclaracionBienes.objects.filter(filtro).values_list('persona_id', flat=True)
                    list_ids = []
                    for id in ids_personas:
                        person = Persona.objects.get(pk=id)
                        if person.activo():
                            list_ids.append(id)
                    if estado == '1':
                        filtro = filtro & Q(persona_id__in=list_ids)
                    else:
                        filtro = filtro & ~Q(persona_id__in=list_ids)

                declaraciones = DeclaracionBienes.objects.filter(filtro).order_by('fecha_creacion', 'persona__cedula', 'tipodeclaracion')

                row_num = 2
                for d in declaraciones:
                    url_archivo = dominio_sistema_base(request)  + d.archivo.url if d.archivo else ''
                    ws.write(row_num, 0, d.fecha_creacion.strftime('%d-%m-%Y') if d.fecha_creacion else '', font_style_center)
                    ws.write(row_num, 1, d.persona.cedula, font_style_center)
                    ws.write(row_num, 2, d.persona.nombre_completo(), font_style2)
                    ws.write(row_num, 3, d.get_tipodeclaracion_display(), font_style_center)
                    ws.write(row_num, 4, d.fecha.strftime('%d-%m-%Y') if d.fecha else '', font_style_center)
                    ws.write(row_num, 5, d.fechaperiodoinicio.strftime('%d-%m-%Y') if d.fechaperiodoinicio else '', font_style_center)
                    ws.write(row_num, 6, d.codigobarra, font_style_center)
                    ws.write(row_num, 7, d.departamento.nombre if d.departamento else '', font_style2)
                    ws.write(row_num, 8, d.denominacionpuesto.descripcion if d.denominacionpuesto else '', font_style2)
                    ws.write(row_num, 9, url_archivo, font_style2)
                    ws.write(row_num, 10, 'ACTIVO' if d.persona.activo() else 'INACTIVO', font_style_center)
                    row_num += 1

                filename = generar_nombre("reporte_declaraciones_", '.xls') + '.xls'
                directory_path = os.path.join(MEDIA_ROOT, 'reportes', 'reporte_declaraciones')
                os.makedirs(directory_path, exist_ok=True)
                wb.save(os.path.join(directory_path, filename))
                url_archivo = dominio_sistema_base(request) + MEDIA_URL + 'reportes/reporte_declaraciones/' + filename
                return JsonResponse({"result": False, "to": url_archivo})
            except Exception as ex:
                return JsonResponse({"result": True, "mensaje": u"Error al obtener los datos." + ex.__str__()})


        # Acción de personal

        elif action == 'addaccion':
            try:
                f = PersonaAccionesForm(request.POST, request.FILES)
                personaaccion = Persona.objects.get(pk=encrypt_id(request.POST['id']))
                if f.is_valid():
                    newfile = None
                    if 'archivo' in request.FILES:
                        newfile = request.FILES['archivo']
                        newfile._name = generar_nombre("accion_", newfile._name)
                        exte = ext_archive(newfile._name)
                        if newfile.size > 5242880:
                            return JsonResponse({"result": "bad", "mensaje": u"Error, archivo mayor a 5 Mb."})
                        if not exte == '.pdf':
                            return JsonResponse({"result": "bad", "mensaje": u"Solo archivos .pdf"})
                    acciones = AccionPersonal(persona=personaaccion,
                                              numerodocumento=f.cleaned_data['numerodocumento'],
                                              tipo=f.cleaned_data['tipo'],
                                              motivoaccion=f.cleaned_data['motivoaccion'],
                                              denominacionpuesto=f.cleaned_data['denominacionpuesto'],
                                              departamento=f.cleaned_data['unidadorganica'],
                                              rmu=f.cleaned_data['remuneracion'],
                                              explicacion=f.cleaned_data['explicacion'],
                                              fechadesde=f.cleaned_data['fechainicio'],
                                              documento=f.cleaned_data['documento'],
                                              estadoarchivo=4,
                                              finalizado=True,
                                              archivo=newfile)
                    acciones.save(request)
                    log(u'Adiciono accion de personal: %s [%s]' % (acciones, acciones.id), request, "add")
                    return JsonResponse({'result': False, 'mensaje': 'Guardado con exito'})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': f'Error al guardar los datos {ex}'})

        elif action == 'editaccionpersonal':

            try:

                f = PersonaAccionesForm(request.POST, request.FILES)

                accion = AccionPersonal.objects.get(pk=request.POST['id'])

                if f.is_valid():

                    newfile = None

                    if 'archivo' in request.FILES:

                        arch = request.FILES['archivo']

                        extencion = arch._name.split('.')

                        exte = extencion[1]

                        if arch.size > 5242880:
                            return JsonResponse({"result": "bad", "mensaje": u"Error, archivo mayor a 5 Mb."})

                        if not exte == 'pdf':
                            return JsonResponse({"result": "bad", "mensaje": u"Solo archivos .pdf"})

                        newfile = request.FILES['archivo']

                        newfile._name = generar_nombre("contrato_", newfile._name)

                        accion.archivo = newfile

                    accion.numerodocumento = f.cleaned_data['numerodocumento']

                    accion.tipo = f.cleaned_data['tipo']

                    accion.motivoaccion = f.cleaned_data['motivoaccion']

                    accion.denominacionpuesto = f.cleaned_data['denominacionpuesto']

                    accion.departamento = f.cleaned_data['unidadorganica']

                    accion.rmu = f.cleaned_data['remuneracion']

                    accion.explicacion = f.cleaned_data['explicacion']

                    accion.fechadesde = f.cleaned_data['fechainicio']

                    accion.documento = f.cleaned_data['documento']

                    accion.save(request)

                    log(u'Edito la accion: %s [%s]' % (accion.persona, accion.id), request, "edit")

                    return JsonResponse({'result': False, 'mensaje': 'Guardado con exito'})

                else:

                    raise NameError('Error')

            except Exception as ex:

                transaction.set_rollback(True)

                return JsonResponse({'result': True, 'mensaje': f'Error al guardar los datos {ex}'})

        elif action == 'editaccionpersonal_old':

            try:

                f = PersonaAccionesOldForm(request.POST, request.FILES)

                personaacciones = PersonaAcciones.objects.get(pk=request.POST['id'])

                if 'archivo' in request.FILES:

                    arch = request.FILES['archivo']

                    extencion = arch._name.split('.')

                    exte = extencion[1]

                    if arch.size > 5242880:
                        return JsonResponse({"result": "bad", "mensaje": u"Error, archivo mayor a 5 Mb."})

                    if not exte == 'pdf':
                        return JsonResponse({"result": "bad", "mensaje": u"Solo archivos .pdf"})

                if f.is_valid():

                    newfile = None

                    accion_new = AccionPersonal.objects.filter(persona=personaacciones.persona,

                                                               numerodocumento=personaacciones.numerodocumento,

                                                               tipo=personaacciones.tipo, status=True).first()

                    personaacciones.numerodocumento = f.cleaned_data['numerodocumento']

                    personaacciones.tipo = f.cleaned_data['tipo']

                    personaacciones.motivo = f.cleaned_data['motivo']

                    personaacciones.remuneracion = f.cleaned_data['remuneracion']

                    personaacciones.explicacion = f.cleaned_data['explicacion']

                    personaacciones.fecharige = f.cleaned_data['fecharige']

                    personaacciones.ubicacionfisico = f.cleaned_data['ubicacionfisico']

                    if 'archivo' in request.FILES:
                        newfile = request.FILES['archivo']

                        newfile._name = generar_nombre("accion_", newfile._name)

                        personaacciones.archivo = newfile

                    personaacciones.save(request)

                    if not accion_new:

                        accion_new = AccionPersonal(persona=personaacciones.persona,

                                                    numerodocumento=personaacciones.numerodocumento,

                                                    tipo=personaacciones.tipo,

                                                    motivoaccion=personaacciones.motivo,

                                                    denominacionpuesto=f.cleaned_data['denominacionpuesto'],

                                                    departamento=f.cleaned_data['unidadorganica'],

                                                    rmu=personaacciones.remuneracion,

                                                    explicacion=personaacciones.explicacion,

                                                    fechadesde=personaacciones.fecharige,

                                                    documento=personaacciones.ubicacionfisico,

                                                    estadoarchivo=4,

                                                    finalizado=True,

                                                    archivo=personaacciones.archivo)

                    else:

                        accion_new.numerodocumento = personaacciones.numerodocumento,

                        accion_new.tipo = personaacciones.tipo,

                        accion_new.motivoaccion = personaacciones.motivo,

                        accion_new.denominacionpuesto = f.cleaned_data['denominacionpuesto'],

                        accion_new.departamento = f.cleaned_data['unidadorganica'],

                        accion_new.rmu = personaacciones.remuneracion,

                        accion_new.explicacion = personaacciones.explicacion,

                        accion_new.fechadesde = personaacciones.fecharige,

                        accion_new.documento = personaacciones.ubicacionfisico,

                        accion_new.estadoarchivo = 4,

                        accion_new.finalizado = True,

                        accion_new.archivo = personaacciones.archivo

                    accion_new.save(request)

                    personaacciones.migrado = True

                    personaacciones.save(request)

                    log(u'Edito accion de personal: %s [%s]' % (accion_new.persona, accion_new.id), request, "edit")

                    return JsonResponse({'result': False, 'mensaje': 'Guardado con exito'})

                else:

                    raise NameError('Error')

            except Exception as ex:

                transaction.set_rollback(True)

                return JsonResponse({'result': True, 'mensaje': f'Error al guardar los datos {ex}'})

        elif action == 'delaccionpersonal':

            try:

                accion = AccionPersonal.objects.get(pk=encrypt(request.POST['id']))

                accion.status = False

                accion.save(request)

                res_json = {"error": False}

            except Exception as ex:

                transaction.set_rollback(True)

                res_json = {'error': True, "message": "Error: {}".format(ex)}

            return JsonResponse(res_json, safe=False)

        elif action == 'delaccionpersonal_old':

            try:

                accion = PersonaAcciones.objects.get(pk=encrypt(request.POST['id']))

                accion.status = False

                accion.save(request)

                res_json = {"error": False}

            except Exception as ex:

                transaction.set_rollback(True)

                res_json = {'error': True, "message": "Error: {}".format(ex)}

            return JsonResponse(res_json, safe=False)

        elif action == 'verificaraccion':
            with transaction.atomic():
                try:
                    val = eval(request.POST['val'].capitalize())
                    registro = PersonaAcciones.objects.get(pk=int(request.POST['id']))
                    registro.estado = val
                    registro.save(request)
                    log(u'Verifico accion de personal: %s - %s' % (registro,val), request, "edit")
                    return JsonResponse({"result": True, 'mensaje': 'Cambios guardados'})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": False})

        # Contratos

        elif action == 'addcontrato':
            try:
                f = PersonaContratosForm(request.POST, request.FILES)
                personacontrato = Persona.objects.get(pk=int(request.POST['id']))
                if f.is_valid():
                    newfile = None
                    if 'archivo' in request.FILES:
                        arch = request.FILES['archivo']
                        extencion = arch._name.split('.')
                        exte = extencion[1]
                        if arch.size > 5242880:
                            return JsonResponse({"result": "bad", "mensaje": u"Error, archivo mayor a 5 Mb."})
                        if not exte == 'pdf':
                            return JsonResponse({"result": "bad", "mensaje": u"Solo archivos .pdf"})
                        newfile = request.FILES['archivo']
                        newfile._name = generar_nombre("contrato_", newfile._name)
                    contratos = PersonaContratos(persona=personacontrato,
                                                 numerodocumento=f.cleaned_data['numerodocumento'],
                                                 contratacionrelacionada=f.cleaned_data[
                                                     'contratacionrelacionada'],
                                                 relacionies=f.cleaned_data['relacionies'],
                                                 denominacionpuesto=f.cleaned_data['denominacionpuesto'],
                                                 unidadorganica=f.cleaned_data['unidadorganica'],
                                                 remuneracion=f.cleaned_data['remuneracion'],
                                                 explicacion=f.cleaned_data['explicacion'],
                                                 fechainicio=f.cleaned_data['fechainicio'],
                                                 fechafin=f.cleaned_data['fechafin'],
                                                 estado=False,
                                                 regimenlaboral=f.cleaned_data['regimenlaboral'],
                                                 dedicacionlaboral=f.cleaned_data['dedicacionlaboral'],
                                                 archivo=newfile)
                    contratos.save(request)
                    log(u'Adiciono persona a contrato: %s [%s]' % (contratos, contratos.id), request, "add")
                    return JsonResponse({'result': False, 'mensaje': 'Guardado con exito'})
                else:
                    raise NameError('Error')
            except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, 'mensaje': f'Error al guardar los datos {ex}'})

        elif action == 'descargarcontratos':
            try:
                f = DescargarContratosForm(request.POST, request.FILES)
                if f.is_valid():
                    fechai = f.cleaned_data['fechainicio']
                    fechaf = f.cleaned_data['fechafin']
                    persona = Persona.objects.get(pk=int(request.POST['id']))
                    filtro = Q(status=True)
                    fechadesdestr = ''
                    fechahastastr = ''
                    if fechai:
                        filtro = filtro & Q(fechainicio__gte=fechai)
                        fechadesdestr = fechai.strftime('%d-%m-%Y')
                    if fechaf:
                        filtro = filtro & Q(fechafin__lte=fechaf)
                        fechahastastr = fechaf.strftime('%d-%m-%Y')
                    contratos = persona.personacontratos_set.filter(filtro).order_by("-fechainicio")

                    directory_p = os.path.join(MEDIA_ROOT, 'talento_humano')
                    directory = os.path.join(directory_p, 'docs_comprimidos')
                    os.makedirs(directory_p, exist_ok=True)
                    os.makedirs(directory, exist_ok=True)
                    name_zip = generar_nombre(persona.nombre_completo_minus(), '') + "_" + fechadesdestr + "_" + fechahastastr + '.zip'
                    url = os.path.join(SITE_STORAGE, 'media', 'talento_humano', 'docs_comprimidos', name_zip)
                    fantasy_zip = zipfile.ZipFile(url, 'w')

                    carpeta = f'{unidecode(persona.nombre_completo_minus())}_{fechadesdestr}_{fechahastastr}'
                    for index, contrato in enumerate(contratos):
                        if contrato.archivo:
                            name_file = unidecode(f'Contrato_{contrato.fechainicio.strftime("%d-%m-%Y")}_{contrato.fechafin.strftime("%d-%m-%Y")}')
                            url_file = contrato.archivo.url
                            ext = url_file[url_file.rfind("."):].lower()
                            ruta_archivo_zip = os.path.join(carpeta, f'{index + 1}.{name_file}{ext}')
                            fantasy_zip.write(contrato.archivo.path, ruta_archivo_zip)

                    fantasy_zip.close()
                    url = f"{MEDIA_URL}talento_humano/docs_comprimidos/{name_zip}"
                    url_archivo_zip = dominio_sistema + url
                    return JsonResponse({'result': False, 'mensaje': 'Guardado con exito', 'to': url_archivo_zip})
                else:
                    raise NameError('Error')
            except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, 'mensaje': f'Error al descargar archivos {ex}'})

        elif action == 'comprimidofotospersonal':
            try:
                noti = Notificacion(cuerpo='Se esta procesando la descarga de fotos de perfil',
                                    titulo='Archivo .zip en proceso',
                                    destinatario=persona,
                                    url='/notificacion',
                                    prioridad=1, app_label='sga-sagest',
                                    fecha_hora_visible=datetime.now() + timedelta(days=1), tipo=2,
                                    en_proceso=True)
                noti.save(request)
                descargar_comprimidos_fotos_perfi_jpg_background(request=request, data=data, notif=noti.pk).start()

                return JsonResponse({'result': 'ok', 'mensaje': 'Se esta procesando la descarga de fotos de perfil'})
            except Exception as ex:
                return JsonResponse({'result': 'bad', 'mensaje': f'Error al descargar archivos {ex}'})

        elif action == 'editcontrato':
            try:
                f = PersonaContratosForm(request.POST, request.FILES)
                personacontrato = PersonaContratos.objects.get(pk=request.POST['id'])
                if f.is_valid():
                    newfile = None
                    if 'archivo' in request.FILES:
                        arch = request.FILES['archivo']
                        extencion = arch._name.split('.')
                        exte = extencion[1]
                        if arch.size > 5242880:
                            return JsonResponse({"result": "bad", "mensaje": u"Error, archivo mayor a 5 Mb."})
                        if not exte == 'pdf':
                            return JsonResponse({"result": "bad", "mensaje": u"Solo archivos .pdf"})
                        newfile = request.FILES['archivo']
                        newfile._name = generar_nombre("contrato_", newfile._name)
                        personacontrato.archivo = newfile
                    personacontrato.numerodocumento=f.cleaned_data['numerodocumento']
                    personacontrato.contratacionrelacionada=f.cleaned_data['contratacionrelacionada']
                    personacontrato.relacionies=f.cleaned_data['relacionies']
                    personacontrato.denominacionpuesto=f.cleaned_data['denominacionpuesto']
                    personacontrato.unidadorganica=f.cleaned_data['unidadorganica']
                    personacontrato.remuneracion=f.cleaned_data['remuneracion']
                    personacontrato.explicacion=f.cleaned_data['explicacion']
                    personacontrato.fechainicio=f.cleaned_data['fechainicio']
                    personacontrato.fechafin=f.cleaned_data['fechafin']
                    personacontrato.regimenlaboral = f.cleaned_data['regimenlaboral']
                    personacontrato.dedicacionlaboral = f.cleaned_data['dedicacionlaboral']
                    personacontrato.save(request)
                    log(u'Edito el contrato: %s [%s]' % (personacontrato.persona,personacontrato.id), request, "edit")
                    return JsonResponse({'result': False, 'mensaje': 'Guardado con exito'})
                else:
                     raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': f'Error al guardar los datos {ex}'})

        elif action == 'delcontrato':
            try:
                eliminarcontrato = PersonaContratos.objects.get(pk=encrypt(request.POST['id']))
                eliminarcontrato.delete()
                res_json = {"error": False}
            except Exception as ex:
                transaction.set_rollback(True)
                res_json = {'error': True, "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)

        elif action == 'verificarcontrato':
            with transaction.atomic():
                try:
                    val = eval(request.POST['val'].capitalize())
                    registro = PersonaContratos.objects.get(pk=int(request.POST['id']))
                    registro.estado = val
                    registro.save(request)
                    log(u'Verifico contrato: %s - %s' % (registro,val), request, "edit")
                    return JsonResponse({"result": True, 'mensaje': 'Cambios guardados'})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": False})


        # Experiencia laboral

        elif action == 'addexperiencia':
            try:
                personaexperiencia = Persona.objects.get(id=request.POST['id'])
                form = ExperienciaLaboralForm(request.POST, request.FILES)
                if 'archivo' in request.FILES:
                    d = request.FILES['archivo']
                    if d.size > 4194304:
                        return JsonResponse({"result": "bad", "mensaje": u"Error, archivo mayor a 4 Mb."})
                    else:
                        newfiles = request.FILES['archivo']
                        newfilesd = newfiles._name
                        ext = newfilesd[newfilesd.rfind("."):]
                        if not (ext == '.pdf' or ext == '.png' or ext == '.jpg' or ext == '.jpeg'):
                            return JsonResponse(
                                {"result": "bad", "mensaje": u"Error, solo archivos .pdf, png, jpg, jpeg."})
                if not form.is_valid():
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()],
                                         "mensaje": "Error en el formulario"})
                motivosalida = None
                fechafin = None
                if not form.cleaned_data['vigente']:
                    motivosalida = form.cleaned_data['motivosalida']
                    fechafin = form.cleaned_data['fechafin']
                experiencialaboral = ExperienciaLaboral(persona=personaexperiencia,
                                                        tipoinstitucion=form.cleaned_data['tipoinstitucion'],
                                                        institucion=form.cleaned_data['institucion'],
                                                        cargo=form.cleaned_data['cargo'],
                                                        departamento=form.cleaned_data['departamento'],
                                                        pais=form.cleaned_data['pais'],
                                                        provincia=form.cleaned_data['provincia'],
                                                        canton=form.cleaned_data['canton'],
                                                        parroquia=form.cleaned_data['parroquia'],
                                                        fechainicio=form.cleaned_data['fechainicio'],
                                                        fechafin=fechafin,
                                                        motivosalida=motivosalida,
                                                        regimenlaboral=form.cleaned_data['regimenlaboral'],
                                                        horassemanales=form.cleaned_data['horassemanales'],
                                                        dedicacionlaboral=form.cleaned_data['dedicacionlaboral'],
                                                        actividadlaboral=form.cleaned_data['actividadlaboral'],
                                                        observaciones=form.cleaned_data['observaciones'])
                experiencialaboral.save(request)
                if 'archivo' in request.FILES:
                    newfile = request.FILES['archivo']
                    newfile._name = generar_nombre("experiencialaboral_", newfile._name)
                    experiencialaboral.archivo = newfile
                    experiencialaboral.save(request)
                log(u'Adiciono experiencia laboral: %s' % experiencialaboral, request, "add")
                return JsonResponse({'result': False, 'mensaje': 'Guardado con exito'})
            except Exception as ex:
                msg = ex.__str__()
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': f'Error al guardar los datos {ex}'})

        elif action == 'editexperiencia':
            try:
                persona = request.session['persona']
                form = ExperienciaLaboralForm(request.POST, request.FILES)
                if 'archivo' in request.FILES:
                    d = request.FILES['archivo']
                    if d.size > 4194304:
                        return JsonResponse({"result": "bad", "mensaje": u"Error, archivo mayor a 4 Mb."})
                    else:
                        newfiles = request.FILES['archivo']
                        newfilesd = newfiles._name
                        ext = newfilesd[newfilesd.rfind("."):]
                        if not (ext == '.pdf' or ext == '.png' or ext == '.jpg' or ext == '.jpeg'):
                            return JsonResponse(
                                {"result": "bad", "mensaje": u"Error, solo archivos .pdf, png, jpg, jpeg."})
                if not form.is_valid():
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()],
                                         "mensaje": "Error en el formulario"})
                motivosalida = None
                fechafin = None
                if not form.cleaned_data['vigente']:
                    motivosalida = form.cleaned_data['motivosalida']
                    fechafin = form.cleaned_data['fechafin']
                experiencialaboral = ExperienciaLaboral.objects.get(pk=encrypt_id(request.POST['id']))
                experiencialaboral.tipoinstitucion = form.cleaned_data['tipoinstitucion']
                experiencialaboral.institucion = form.cleaned_data['institucion']
                experiencialaboral.cargo = form.cleaned_data['cargo']
                experiencialaboral.departamento = form.cleaned_data['departamento']
                experiencialaboral.pais = form.cleaned_data['pais']
                experiencialaboral.provincia = form.cleaned_data['provincia']
                experiencialaboral.canton = form.cleaned_data['canton']
                experiencialaboral.parroquia = form.cleaned_data['parroquia']
                experiencialaboral.fechainicio = form.cleaned_data['fechainicio']
                experiencialaboral.fechafin = fechafin
                experiencialaboral.motivosalida = motivosalida
                experiencialaboral.regimenlaboral = form.cleaned_data['regimenlaboral']
                experiencialaboral.horassemanales = form.cleaned_data['horassemanales']
                experiencialaboral.dedicacionlaboral = form.cleaned_data['dedicacionlaboral']
                experiencialaboral.actividadlaboral = form.cleaned_data['actividadlaboral']
                experiencialaboral.observaciones = form.cleaned_data['observaciones']
                experiencialaboral.save(request)
                if 'archivo' in request.FILES:
                    newfile = request.FILES['archivo']
                    newfile._name = generar_nombre("experiencialaboral_", newfile._name)
                    experiencialaboral.archivo = newfile
                    experiencialaboral.save(request)
                log(u'Modifico experiencia laboral: %s' % experiencialaboral, request, "edit")
                return JsonResponse({'result': False, 'mensaje': 'Guardado con exito'})
            except Exception as ex:
                msg = ex.__str__()
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': f'Error al guardar los datos {ex}'})

        elif action == 'delexperiencia':
            try:
                experiencia = ExperienciaLaboral.objects.get(pk=encrypt_id(request.POST['id']))
                log(u"Elimino experiencia laboral: %s" % experiencia, request, "del")
                experiencia.delete()
                res_js = {'error': False}
            except Exception as ex:
                transaction.set_rollback(True)
                err = 'Error on line {}'.format(sys.exc_info()[-1].tb_lineno)
                msg_err = f"Ocurrio un error: {ex.__str__()}. {err}"
                res_js = {'error': True, 'mensaje': msg_err}
            return JsonResponse(res_js)

        elif action == 'verificarexperiencia':
            with transaction.atomic():
                try:
                    val = eval(request.POST['val'].capitalize())
                    registro = ExperienciaLaboral.objects.get(pk=int(request.POST['id']))
                    registro.verificado = val
                    registro.save(request)
                    log(u'Verifico experiencia laboral: %s - %s' % (registro,val), request, "edit")
                    return JsonResponse({"result": True, 'mensaje': 'Cambios guardados'})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": False})


        # Datos institucionales
        elif action == 'editdatosinstitucion':
            try:
                form = DatosPersonaInstitucionForm(request.POST)
                if form.is_valid():
                    personaadmin = Persona.objects.get(pk=int(request.POST['id']))
                    personaadmin.identificacioninstitucion = form.cleaned_data['indicebiometrico']
                    personaadmin.regitrocertificacion = form.cleaned_data['registro']
                    personaadmin.servidorcarrera = form.cleaned_data['servidorcarrera']
                    personaadmin.telefonoextension = form.cleaned_data['extension']
                    personaadmin.emailinst = form.cleaned_data['correoinstitucional']
                    personaadmin.fechaingresoies = form.cleaned_data['fechaingresoies']
                    personaadmin.fechasalidaies = form.cleaned_data['fechasalidaies']
                    personaadmin.concursomeritos = form.cleaned_data['concursomeritos']
                    personaadmin.save(request)
                    log(u'Modifico datos laborales de: %s' % personaadmin, request, "add")
                    return JsonResponse({'result': False, 'mensaje': 'Guardado con exito'})
                else:
                    raise NameError('Error')
            except Exception as ex:
                msg = ex.__str__()
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': f'Error al guardar los datos {ex}'})

    #DATOS PERSONALES

        # Datos personales

        elif action == 'editdatospersonales':
            try:
                administrativo = Persona.objects.get(id=request.POST['id'])
                f = DatosPersonalesForm(request.POST, request.FILES)
                if not f.is_valid():
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()],
                                         "mensaje": "Error en el formulario"})
                if 'archivocedula' in request.FILES:
                    arch = request.FILES['archivocedula']
                    extension = arch._name.split('.')
                    tam = len(extension)
                    exte = extension[tam - 1]
                    if arch.size > 4194304:
                        raise NameError("Error, el tamaño del archivo de cédula es mayor a 4 Mb.")
                    if not exte.lower() == 'pdf':
                        raise NameError("Solo se permiten archivos .pdf")
                if 'archivoraza' in request.FILES:
                    arch_r = request.FILES['archivoraza']
                    extension = arch_r._name.split('.')
                    tam = len(extension)
                    exte = extension[tam - 1]
                    if arch_r.size > 4194304:
                        raise NameError("Error, el tamaño del archivo de etnia es mayor a 4 Mb.")
                    if not exte.lower() in ['pdf']:
                        raise NameError("Solo se permiten archivos .pdf")
                administrativo.pasaporte = f.cleaned_data['pasaporte']
                administrativo.anioresidencia = f.cleaned_data['anioresidencia']
                administrativo.nacimiento = f.cleaned_data['nacimiento']
                administrativo.telefonoextension = f.cleaned_data['extension']
                administrativo.sexo = f.cleaned_data['sexo']
                administrativo.lgtbi = f.cleaned_data['lgtbi']
                administrativo.email = f.cleaned_data['email']
                administrativo.libretamilitar = f.cleaned_data['libretamilitar']
                administrativo.eszurdo = f.cleaned_data['eszurdo']
                administrativo.save(request)
                administrativoext = administrativo.datos_extension()
                administrativoext.estadocivil = f.cleaned_data['estadocivil']
                administrativoext.save(request)

                if 'archivocedula' in request.FILES:
                    newfile = request.FILES['archivocedula']
                    newfile._name = generar_nombre("cedula", newfile._name)

                    documento = administrativo.documentos_personales()
                    if documento is None:
                        documento = PersonaDocumentoPersonal(persona=administrativo,
                                                             cedula=newfile,
                                                             estadocedula=1
                                                             )
                    else:
                        documento.cedula = newfile
                        documento.estadocedula = 1

                    documento.save(request)

                if 'papeleta' in request.FILES:
                    newfile = request.FILES['papeleta']
                    newfile._name = generar_nombre("papeleta", newfile._name)

                    documento = administrativo.documentos_personales()
                    if documento is None:
                        documento = PersonaDocumentoPersonal(persona=administrativo,
                                                             papeleta=newfile,
                                                             estadopapeleta=1
                                                             )
                    else:
                        documento.papeleta = newfile
                        documento.estadopapeleta = 1

                    documento.save(request)

                if 'archivolibretamilitar' in request.FILES:
                    newfile = request.FILES['archivolibretamilitar']
                    newfile._name = generar_nombre("libretamilitar", newfile._name)

                    documento = administrativo.documentos_personales()
                    if documento is None:
                        documento = PersonaDocumentoPersonal(persona=administrativo,
                                                             libretamilitar=newfile,
                                                             estadolibretamilitar=1
                                                             )
                    else:
                        documento.libretamilitar = newfile
                        documento.estadolibretamilitar = 1

                    documento.save(request)

                perfil = administrativo.mi_perfil()
                perfil.raza = f.cleaned_data['raza']
                perfil.nacionalidadindigena = f.cleaned_data['nacionalidadindigena']
                if 'archivoraza' in request.FILES:
                    arch_r._name = generar_nombre("archivoraza", arch_r._name)
                    perfil.archivoraza = arch_r
                    perfil.estadoarchivoraza = 1
                else:
                    if perfil.raza.id != 1:
                        perfil.archivoraza = None
                        perfil.estadoarchivoraza = None
                perfil.save(request)
                log(u'Modifico datos personales: %s' % administrativo, request, "edit")
                return JsonResponse({'result': False, 'mensaje': 'Guardado con exito'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': f'Error al guardar los datos: {ex}'})

        elif action == 'editdatosnacimiento':
            try:
                administrativo = Persona.objects.get(id=request.POST['id'])
                f = DatosNacimientoForm(request.POST)
                if not f.is_valid():
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()],
                                         "mensaje": "Error en el formulario"})
                administrativo.paisnacimiento = f.cleaned_data['paisnacimiento']
                administrativo.provincianacimiento = f.cleaned_data['provincianacimiento']
                administrativo.cantonnacimiento = f.cleaned_data['cantonnacimiento']
                administrativo.parroquianacimiento = f.cleaned_data['parroquianacimiento']
                administrativo.paisnacionalidad = f.cleaned_data['paisnacionalidad']
                administrativo.save(request)
                log(u'Modifico datos de nacimiento: %s' % administrativo, request, "edit")
                return JsonResponse({'result': False, 'mensaje': 'Guardado con exito'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': f'Error al guardar los datos {ex}'})

        elif action == 'editdatosdomicilio':
            try:
                if 'archivocroquis' in request.FILES:
                    newfile = request.FILES['archivocroquis']
                    if newfile.size > 2194304:
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"Error, Tamaño de archivo Maximo permitido es de 2Mb"})

                if 'archivoplanillaluz' in request.FILES:
                    newfile = request.FILES['archivoplanillaluz']
                    if newfile.size > 2194304:
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"Error, Tamaño de archivo Maximo permitido es de 2Mb"})

                if 'serviciosbasico' in request.FILES:
                    newfile2 = request.FILES['serviciosbasico']
                    if newfile2.size > 2194304:
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"Error, Tamaño de archivo Maximo permitido es de 2Mb"})

                administrativo = Persona.objects.get(id=request.POST['id'])
                documento_p=administrativo.documentos_personales()
                f = DatosDomicilioForm(request.POST)
                if 'pais' in request.POST and request.POST['pais'] and int(request.POST['pais']) == 1:
                    if 'provincia' in request.POST and not request.POST['provincia']:
                        raise NameError('Debe ingresa una provincia')
                    if 'canton' in request.POST and not request.POST['canton']:
                        raise NameError('Debe ingresa una canton')
                    if 'parroquia' in request.POST and not request.POST['parroquia']:
                        raise NameError('Debe ingresa una parroquia')
                if not f.is_valid():
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()],
                                         "mensaje": "Error en el formulario"})

                newfile = None
                administrativo.pais = f.cleaned_data['pais']
                administrativo.provincia = f.cleaned_data['provincia']
                administrativo.canton = f.cleaned_data['canton']
                administrativo.sector = f.cleaned_data['sector']
                administrativo.parroquia = f.cleaned_data['parroquia']
                administrativo.direccion = f.cleaned_data['direccion']
                administrativo.direccion2 = f.cleaned_data['direccion2']
                administrativo.ciudadela = f.cleaned_data['ciudadela']
                administrativo.num_direccion = f.cleaned_data['num_direccion']
                administrativo.telefono_conv = f.cleaned_data['telefono_conv']
                administrativo.telefono = f.cleaned_data['telefono']
                administrativo.tipocelular = f.cleaned_data['tipocelular']
                administrativo.referencia = f.cleaned_data['referencia']
                administrativo.zona = int(f.cleaned_data['zona'])
                administrativo.save(request)
                if 'archivocroquis' in request.FILES:
                    newfile = request.FILES['archivocroquis']
                    newfile._name = generar_nombre("croquis_", newfile._name)
                    administrativo.archivocroquis = newfile
                    administrativo.save(request)

                if 'archivoplanillaluz' in request.FILES:
                    newfile = request.FILES['archivoplanillaluz']
                    newfile._name = generar_nombre("planilla_luz_", newfile._name)
                    administrativo.archivoplanillaluz = newfile
                    administrativo.save(request)
                if 'serviciosbasico' in request.FILES:
                    newfile = request.FILES['serviciosbasico']
                    newfile._name = generar_nombre("serviciobasico_", newfile._name)
                    documento_p.serviciosbasico = newfile
                    documento_p.estadoserviciosbasico = 1
                    documento_p.save(request)

                log(u'Modifico datos de domicilio: %s' % administrativo, request, "edit")
                return JsonResponse({'result': False,'mensaje':'Guardado con éxito'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': f'Error al guardar los datos {ex}'})

        # Datos familiares

        elif action == 'addfamiliar':
            try:
                administrativo = Persona.objects.get(id=request.POST['id'])
                f = FamiliarForm(request.POST)
                if f.is_valid():
                    edit_d = eval(request.POST.get('edit_d', ''))
                    cedula = f.cleaned_data['identificacion'].strip()
                    if administrativo.personadatosfamiliares_set.filter(identificacion=f.cleaned_data['identificacion'], status=True).exists():
                        raise NameError('El familiar se encuentra registrado.')
                    nombres = f"{f.cleaned_data['apellido1']} {f.cleaned_data['apellido2']} {f.cleaned_data['nombre']}"
                    familiar = PersonaDatosFamiliares(persona=administrativo,
                                                      identificacion=cedula,
                                                      nombre=nombres,
                                                      fallecido=f.cleaned_data['fallecido'],
                                                      nacimiento=f.cleaned_data['nacimiento'],
                                                      parentesco=f.cleaned_data['parentesco'],
                                                      tienediscapacidad=f.cleaned_data['tienediscapacidad'],
                                                      telefono=f.cleaned_data['telefono'],
                                                      telefono_conv=f.cleaned_data['telefono_conv'],
                                                      niveltitulacion=f.cleaned_data['niveltitulacion'],
                                                      ingresomensual=f.cleaned_data['ingresomensual'],
                                                      formatrabajo=f.cleaned_data['formatrabajo'],
                                                      trabajo=f.cleaned_data['trabajo'],
                                                      convive=f.cleaned_data['convive'],
                                                      sustentohogar=f.cleaned_data['sustentohogar'],
                                                      # rangoedad=f.cleaned_data['rangoedad'],
                                                      essustituto=f.cleaned_data['essustituto'],
                                                      autorizadoministerio=f.cleaned_data['autorizadoministerio'],
                                                      tipodiscapacidad=f.cleaned_data['tipodiscapacidad'],
                                                      porcientodiscapacidad=f.cleaned_data['porcientodiscapacidad'],
                                                      carnetdiscapacidad=f.cleaned_data['carnetdiscapacidad'],
                                                      institucionvalida=f.cleaned_data['institucionvalida'],
                                                      tipoinstitucionlaboral=f.cleaned_data['tipoinstitucionlaboral'],
                                                      negocio=f.cleaned_data['negocio'],
                                                      esservidorpublico=f.cleaned_data['esservidorpublico'],
                                                      bajocustodia=f.cleaned_data['bajocustodia'],
                                                      centrocuidado=f.cleaned_data['centrocuidado'] if f.cleaned_data['centrocuidado'] else 0,
                                                      centrocuidadodesc=f.cleaned_data['centrocuidadodesc'],
                                                      tienenegocio=f.cleaned_data['tienenegocio'])
                    familiar.save(request)
                    if 'cedulaidentidad' in request.FILES:
                        newfile = request.FILES['cedulaidentidad']
                        newfile._name = generar_nombre("cedulaidentidad_", newfile._name)
                        familiar.cedulaidentidad = newfile
                        familiar.save(request)
                    if 'ceduladiscapacidad' in request.FILES:
                        newfile = request.FILES['ceduladiscapacidad']
                        newfile._name = generar_nombre("ceduladiscapacidad_", newfile._name)
                        familiar.ceduladiscapacidad = newfile
                        familiar.save(request)
                    if 'archivoautorizado' in request.FILES:
                        newfile = request.FILES['archivoautorizado']
                        newfile._name = generar_nombre("archivoautorizado_", newfile._name)
                        familiar.archivoautorizado = newfile
                        familiar.save(request)
                    if 'cartaconsentimiento' in request.FILES:
                        newfile = request.FILES['cartaconsentimiento']
                        newfile._name = generar_nombre(f"cartaconsentimiento_{administrativo.usuario.username}", newfile._name)
                        familiar.cartaconsentimiento = newfile
                        familiar.save(request)
                    if 'archivocustodia' in request.FILES:
                        newfile = request.FILES['archivocustodia']
                        newfile._name = generar_nombre(f"archivocustodia_{administrativo.usuario.username}", newfile._name)
                        familiar.archivocustodia = newfile
                        familiar.save(request)

                    pers = Persona.objects.filter(Q(pasaporte=cedula) | Q(cedula=cedula) | Q(pasaporte=('VS' + cedula)) | Q(cedula=cedula[2:]),status=True).first()
                    if not pers:
                        pers = Persona(cedula=f.cleaned_data['identificacion'],
                                       nombres=f.cleaned_data['nombre'],
                                       apellido1=f.cleaned_data['apellido1'],
                                       apellido2=f.cleaned_data['apellido2'],
                                       nacimiento=f.cleaned_data['nacimiento'],
                                       telefono=f.cleaned_data['telefono'],
                                       sexo=f.cleaned_data['sexo'],
                                       telefono_conv=f.cleaned_data['telefono_conv'],
                                       )
                        pers.save(request)
                        log(u'Adiciono persona: %s' % administrativo, request, "add")
                    elif len(pers.mis_perfilesusuarios()) == 1 and pers.tiene_usuario_externo():
                        pers.cedula = f.cleaned_data['identificacion']
                        pers.nombres = f.cleaned_data['nombre']
                        pers.apellido1 = f.cleaned_data['apellido1']
                        pers.apellido2 = f.cleaned_data['apellido2']
                        pers.nacimiento = f.cleaned_data['nacimiento']
                        pers.telefono = f.cleaned_data['telefono']
                        pers.sexo = f.cleaned_data['sexo']
                        pers.telefono_conv = f.cleaned_data['telefono_conv']
                        pers.save(request)
                        log(u'Edito familiar de usuario: %s' % pers, request, "edit")
                    if not pers.tiene_perfil():
                        if not Externo.objects.filter(persona=pers, status=True):
                            externo = Externo(persona=pers)
                            externo.save(request)
                            log(u'Adiciono externo: %s' % pers, request, "add")
                        perfil = PerfilUsuario(persona=pers, externo=externo)
                        perfil.save(request)
                        log(u'Adiciono perfil de usuario: %s' % perfil, request, "add")

                    perfil_i=pers.mi_perfil()
                    if not edit_d and perfil_i.tienediscapacidad:
                        familiar.tienediscapacidad=perfil_i.tienediscapacidad
                        familiar.tipodiscapacidad=perfil_i.tipodiscapacidad
                        familiar.porcientodiscapacidad=perfil_i.porcientodiscapacidad
                        familiar.carnetdiscapacidad=perfil_i.carnetdiscapacidad
                        familiar.institucionvalida=perfil_i.institucionvalida
                        familiar.ceduladiscapacidad=perfil_i.archivo.name if perfil_i.archivo else ''
                        familiar.archivoautorizado=perfil_i.archivovaloracion.name if perfil_i.archivovaloracion else ''
                    elif f.cleaned_data['tienediscapacidad']:
                        perfil_i.tienediscapacidad = f.cleaned_data['tienediscapacidad']
                        perfil_i.tipodiscapacidad = f.cleaned_data['tipodiscapacidad']
                        perfil_i.porcientodiscapacidad = f.cleaned_data['porcientodiscapacidad'] if f.cleaned_data['porcientodiscapacidad'] else 0
                        perfil_i.carnetdiscapacidad = f.cleaned_data['carnetdiscapacidad']
                        perfil_i.institucionvalida = f.cleaned_data['institucionvalida']
                        if 'ceduladiscapacidad' in request.FILES:
                            newfile = request.FILES['ceduladiscapacidad']
                            newfile._name = generar_nombre("archivosdiscapacidad_", newfile._name)
                            perfil_i.archivo = newfile
                            perfil_i.estadoarchivodiscapacidad = 1
                        if 'archivoautorizado' in request.FILES:
                            newfile = request.FILES['archivoautorizado']
                            newfile._name = generar_nombre("archivovaloracionmedica_", newfile._name)
                            perfil_i.archivovaloracion = newfile
                        perfil_i.save(request)
                    if f.cleaned_data['parentesco'].id in [14,11] and not administrativo.apellido1 in [pers.apellido1,pers.apellido2]:
                        familiar.aprobado=False
                    if f.cleaned_data['parentesco'].id == 13 and not pers.personadatosfamiliares_set.filter(personafamiliar=administrativo, status=True).exists():
                        fam_ = PersonaDatosFamiliares(persona=pers,
                                                      personafamiliar=administrativo,
                                                      identificacion=administrativo.cedula,
                                                      nombre=administrativo.nombre_completo_inverso(),
                                                      nacimiento=administrativo.nacimiento,
                                                      parentesco=f.cleaned_data['parentesco'],
                                                      telefono=administrativo.telefono,
                                                      telefono_conv=administrativo.telefono_conv,
                                                      convive=f.cleaned_data['convive'])
                        fam_.save(request)
                        perfil_fam = administrativo.mi_perfil()
                        if perfil_fam.tienediscapacidad:
                            fam_.tienediscapacidad = perfil_fam.tienediscapacidad
                            fam_.tipodiscapacidad = perfil_fam.tipodiscapacidad
                            fam_.porcientodiscapacidad = perfil_fam.porcientodiscapacidad
                            fam_.carnetdiscapacidad = perfil_fam.carnetdiscapacidad
                            fam_.institucionvalida = perfil_fam.institucionvalida
                            fam_.ceduladiscapacidad = perfil_fam.archivo.name if perfil_i.archivo else ''
                            fam_.archivoautorizado = perfil_fam.archivovaloracion.name if perfil_i.archivovaloracion else ''
                        fam_.save(request)
                    familiar.personafamiliar=pers
                    familiar.save(request)
                    if familiar.parentesco.id in [11, 14] or familiar.bajocustodia:
                        per_extension = administrativo.personaextension_set.filter(status=True).first()
                        hijos = per_extension.hijos if per_extension.hijos else 0
                        per_extension.hijos = hijos+1
                        per_extension.save(request)
                    log(u'Adiciono familiar: %s' % familiar, request, "add")
                    return JsonResponse({'result': False, 'mensaje':'Guardado con exito'})
                else:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()],
                                         "mensaje": "Error en el formulario"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', 'mensaje': f'{ex}'})

        elif action == 'editfamiliar':
            try:
                administrativo = Persona.objects.get(id=request.POST['idp'])
                f = FamiliarForm(request.POST)
                f.edit()
                if f.is_valid():
                    familiar = PersonaDatosFamiliares.objects.get(pk=encrypt_id(request.POST['id']))
                    edit_d=eval(request.POST.get('edit_d',''))
                    cedula = f.cleaned_data['identificacion'].strip()
                    if administrativo.personadatosfamiliares_set.filter(identificacion=cedula, status=True).exclude(id=familiar.id).exists():
                        raise NameError(u'El familiar se encuentra registrado.')
                    nombres = f"{f.cleaned_data['apellido1']} {f.cleaned_data['apellido2']} {f.cleaned_data['nombre']}"
                    pers = Persona.objects.filter(Q(pasaporte=cedula) | Q(cedula=cedula) | Q(pasaporte=('VS' + cedula)) | Q(cedula=cedula[2:]),status=True).first()
                    familiar.fallecido = f.cleaned_data['fallecido']
                    familiar.parentesco = f.cleaned_data['parentesco']
                    familiar.tienediscapacidad = f.cleaned_data['tienediscapacidad']
                    familiar.trabajo = f.cleaned_data['trabajo']
                    familiar.niveltitulacion = f.cleaned_data['niveltitulacion']
                    familiar.ingresomensual = f.cleaned_data['ingresomensual']
                    familiar.formatrabajo = f.cleaned_data['formatrabajo']
                    familiar.convive = f.cleaned_data['convive']
                    familiar.sustentohogar = f.cleaned_data['sustentohogar']
                    # familiar.rangoedad = f.cleaned_data['rangoedad']
                    familiar.tienenegocio = f.cleaned_data['tienenegocio']
                    familiar.esservidorpublico = f.cleaned_data['esservidorpublico']
                    familiar.bajocustodia = f.cleaned_data['bajocustodia']
                    familiar.centrocuidado = f.cleaned_data['centrocuidado'] if f.cleaned_data['centrocuidado'] else 0
                    familiar.centrocuidadodesc = f.cleaned_data['centrocuidadodesc']
                    familiar.negocio = ''
                    familiar.tipoinstitucionlaboral = f.cleaned_data['tipoinstitucionlaboral']
                    if familiar.tienenegocio:
                        familiar.negocio = f.cleaned_data['negocio']
                    familiar.save(request)
                    if 'cedulaidentidad' in request.FILES:
                        newfile = request.FILES['cedulaidentidad']
                        newfile._name = generar_nombre("cedulaidentidad_", newfile._name)
                        familiar.cedulaidentidad = newfile
                        familiar.save(request)
                    if 'ceduladiscapacidad' in request.FILES:
                        newfile = request.FILES['ceduladiscapacidad']
                        newfile._name = generar_nombre("ceduladiscapacidad_", newfile._name)
                        familiar.ceduladiscapacidad = newfile
                        familiar.save(request)
                    if 'cartaconsentimiento' in request.FILES:
                        newfile = request.FILES['cartaconsentimiento']
                        newfile._name = generar_nombre(f"cartaconsentimiento_{administrativo.usuario.username}", newfile._name)
                        familiar.cartaconsentimiento = newfile
                        familiar.save(request)
                    if 'archivocustodia' in request.FILES:
                        newfile = request.FILES['archivocustodia']
                        newfile._name = generar_nombre(f"archivocustodia_{administrativo.usuario.username}", newfile._name)
                        familiar.archivocustodia = newfile
                        familiar.save(request)

                    if f.cleaned_data['tienediscapacidad']:
                        if 'archivoautorizado' in request.FILES:
                            newfile = request.FILES['archivoautorizado']
                            newfile._name = generar_nombre("archivoautorizado_", newfile._name)
                            familiar.archivoautorizado = newfile
                            familiar.save(request)
                    else:
                        familiar.archivoautorizado = None
                        familiar.save(request)
                    if not pers:
                        pers = Persona(cedula=f.cleaned_data['identificacion'],
                                       nombres=f.cleaned_data['nombre'],
                                       apellido1=f.cleaned_data['apellido1'],
                                       apellido2=f.cleaned_data['apellido2'],
                                       nacimiento=f.cleaned_data['nacimiento'],
                                       telefono=f.cleaned_data['telefono'],
                                       sexo=f.cleaned_data['sexo'],
                                       telefono_conv=f.cleaned_data['telefono_conv'],
                                       )
                        pers.save(request)
                        log(u'Adiciono persona: %s' % administrativo, request, "add")
                    elif len(pers.mis_perfilesusuarios()) == 1 and pers.tiene_usuario_externo():
                        pers.cedula = cedula
                        pers.nombres = f.cleaned_data['nombre']
                        pers.apellido1 = f.cleaned_data['apellido1']
                        pers.apellido2 = f.cleaned_data['apellido2']
                        pers.nacimiento = f.cleaned_data['nacimiento']
                        pers.telefono = f.cleaned_data['telefono']
                        pers.sexo = f.cleaned_data['sexo']
                        pers.telefono_conv = f.cleaned_data['telefono_conv']
                        pers.save(request)
                        log(u'Edito familiar de usuario: %s' % pers, request, "edit")
                    if not pers.tiene_perfil():
                        if not Externo.objects.filter(persona=pers, status=True):
                            externo = Externo(persona=pers)
                            externo.save(request)
                            log(u'Adiciono externo: %s' % pers, request, "add")
                        perfil = PerfilUsuario(persona=pers, externo=externo)
                        perfil.save(request)
                        log(u'Adiciono perfil de usuario: %s' % perfil, request, "add")
                    if len(pers.mis_perfilesusuarios()) == 1 and pers.tiene_usuario_externo():
                        familiar.identificacion = cedula
                        familiar.nombre = nombres
                        familiar.nacimiento = f.cleaned_data['nacimiento']
                        familiar.telefono = f.cleaned_data['telefono']
                        familiar.telefono_conv = f.cleaned_data['telefono_conv']
                    else:
                        familiar.identificacion = pers.cedula
                        familiar.nombre = pers.nombre_completo_inverso()
                        familiar.nacimiento = pers.nacimiento
                        familiar.telefono = pers.telefono
                        familiar.telefono_conv = pers.telefono_conv
                    perfil_i = pers.mi_perfil()
                    if edit_d:
                        if f.cleaned_data['tienediscapacidad']:
                            familiar.essustituto = f.cleaned_data['essustituto']
                            familiar.autorizadoministerio = f.cleaned_data['autorizadoministerio']
                            familiar.tipodiscapacidad = f.cleaned_data['tipodiscapacidad']
                            familiar.porcientodiscapacidad = f.cleaned_data['porcientodiscapacidad']
                            familiar.carnetdiscapacidad = f.cleaned_data['carnetdiscapacidad']
                            familiar.institucionvalida = f.cleaned_data['institucionvalida']
                            perfil_i.tienediscapacidad = f.cleaned_data['tienediscapacidad']
                            perfil_i.tipodiscapacidad = f.cleaned_data['tipodiscapacidad']
                            perfil_i.porcientodiscapacidad = f.cleaned_data['porcientodiscapacidad'] if f.cleaned_data['porcientodiscapacidad'] else 0
                            perfil_i.carnetdiscapacidad = f.cleaned_data['carnetdiscapacidad']
                            perfil_i.institucionvalida = f.cleaned_data['institucionvalida']
                            if 'ceduladiscapacidad' in request.FILES:
                                newfile = request.FILES['ceduladiscapacidad']
                                newfile._name = generar_nombre("archivosdiscapacidad_", newfile._name)
                                perfil_i.archivo = newfile
                                perfil_i.estadoarchivodiscapacidad = 1
                            if 'archivoautorizado' in request.FILES:
                                newfile = request.FILES['archivoautorizado']
                                newfile._name = generar_nombre("archivovaloracionmedica_", newfile._name)
                                perfil_i.archivovaloracion = newfile
                            perfil_i.save(request)
                        else:
                            familiar.essustituto = False
                            familiar.autorizadoministerio = False
                            familiar.tipodiscapacidad = None
                            familiar.porcientodiscapacidad = None
                            familiar.carnetdiscapacidad = ''
                            familiar.institucionvalida = None
                            perfil_i.tienediscapacidad = False
                            perfil_i.tipodiscapacidad = None
                            perfil_i.porcientodiscapacidad = None
                            perfil_i.carnetdiscapacidad = ''
                            perfil_i.institucionvalida = None
                            perfil_i.save(request)
                    elif perfil_i.tienediscapacidad:
                        familiar.essustituto = f.cleaned_data['essustituto']
                        familiar.autorizadoministerio = f.cleaned_data['autorizadoministerio']
                        familiar.tienediscapacidad = perfil_i.tienediscapacidad
                        familiar.tipodiscapacidad = perfil_i.tipodiscapacidad
                        familiar.porcientodiscapacidad = perfil_i.porcientodiscapacidad
                        familiar.carnetdiscapacidad = perfil_i.carnetdiscapacidad
                        familiar.institucionvalida = perfil_i.institucionvalida
                        familiar.ceduladiscapacidad = perfil_i.archivo.name if perfil_i.archivo else ''
                        familiar.archivoautorizado = perfil_i.archivovaloracion.name if perfil_i.archivovaloracion else ''
                        perfil_i.save(request)
                    if f.cleaned_data['parentesco'].id in [14, 11] and not administrativo.apellido1 in [pers.apellido1, pers.apellido2]:
                        familiar.aprobado = False
                    if f.cleaned_data['parentesco'].id == 13 and not pers.personadatosfamiliares_set.filter(personafamiliar=administrativo, status=True).exists():
                        fam_ = PersonaDatosFamiliares(persona=pers,
                                                      personafamiliar=administrativo,
                                                      identificacion=administrativo.cedula,
                                                      nombre=administrativo.nombre_completo_inverso(),
                                                      nacimiento=administrativo.nacimiento,
                                                      parentesco=f.cleaned_data['parentesco'],
                                                      telefono=administrativo.telefono,
                                                      telefono_conv=administrativo.telefono_conv,
                                                      convive=f.cleaned_data['convive'])
                        fam_.save(request)
                        perfil_fam = administrativo.mi_perfil()
                        if perfil_fam.tienediscapacidad:
                            fam_.tienediscapacidad = perfil_fam.tienediscapacidad
                            fam_.tipodiscapacidad = perfil_fam.tipodiscapacidad
                            fam_.porcientodiscapacidad = perfil_fam.porcientodiscapacidad
                            fam_.carnetdiscapacidad = perfil_fam.carnetdiscapacidad
                            fam_.institucionvalida = perfil_fam.institucionvalida
                            fam_.ceduladiscapacidad = perfil_fam.archivo.name if perfil_i.archivo else ''
                            fam_.archivoautorizado = perfil_fam.archivovaloracion.name if perfil_i.archivovaloracion else ''
                        fam_.save(request)
                    familiar.personafamiliar = pers
                    familiar.save(request)
                    log(u'Modifico familiar: %s' % administrativo, request, "edit")
                    return JsonResponse({'result': False, 'mensaje': 'Guardado con exito'})
                else:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()],
                                         "mensaje": "Error en el formulario"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', 'mensaje': str(ex)})

        elif action == 'delfamiliar':
            try:
                familiar = PersonaDatosFamiliares.objects.get(pk=encrypt_id(request.POST['id']))
                familiar.status = False
                familiar.save(request)
                log(u'Elimino familiar: %s' % familiar, request, "del")
                res_js = {'error': False}
            except Exception as ex:
                transaction.set_rollback(True)
                err = 'Error on line {}'.format(sys.exc_info()[-1].tb_lineno)
                msg_err = f"Ocurrio un error: {ex.__str__()}. {err}"
                res_js = {'error': True, 'mensaje': msg_err}
            return JsonResponse(res_js)

        if action == 'habilitar_familiar':
            try:
                familiar = PersonaDatosFamiliares.objects.get(pk=request.POST['id'])
                familiar.actagenerada=False
                familiar.save()
                log(u"Habilitó carga familiar para modificacion : %s" % familiar, request, "edit")
                return JsonResponse({'result': 'ok'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', 'mensaje': u'Error al habilitar familiar'})


    # DATOS FINANZAS

        # Cuenta bancaria
        elif action == 'addcuentabancaria':
            try:
                f = CuentaBancariaPersonaForm(request.POST)
                administrativo = Persona.objects.get(id=request.POST['id'])
                if f.is_valid():
                    numero = f.cleaned_data['numero'].strip()
                    if administrativo.cuentabancariapersona_set.filter(numero=numero).exists():
                        return JsonResponse({'result': True,
                                             "mensaje": "La cuenta bancaria se encuentra registrada."})
                    cuentabancaria = CuentaBancariaPersona(persona=administrativo,
                                                           numero=f.cleaned_data['numero'],
                                                           estadorevision=1,
                                                           banco=f.cleaned_data['banco'],
                                                           tipocuentabanco=f.cleaned_data['tipocuentabanco'], )
                    cuentabancaria.save(request)
                    if 'archivo' in request.FILES:
                        newfile = request.FILES['archivo']
                        if newfile:
                            newfile._name = generar_nombre("cuentabancaria_", newfile._name)
                            cuentabancaria.archivo = newfile
                            cuentabancaria.save(request)
                        if DistributivoPersona.objects.filter(status=True,persona=administrativo ):
                            lista = ['talento_humano@unemi.edu.ec', 'portizg@unemi.edu.ec']
                            asunto = "Ingresaron nueva cuenta bancaria (archivo)"
                            send_html_mail(asunto, "emails/ingreso_archivos_tthh.html",
                                           {'asunto': asunto, 'd': cuentabancaria.persona.nombre_completo_inverso(),
                                            'fecha': datetime.now().date(), 'escenario': 'cuenta bancaria'}, lista, [],
                                           cuenta=CUENTAS_CORREOS[1][1])
                    log(u'Adiciono cuenta bancaria: %s' % administrativo, request, "add")
                    return JsonResponse({'result': False, 'mensaje': 'Guardado con exito'})
                else:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()],
                                         "mensaje": "Error en el formulario"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': f'{ex}'})

        elif action == 'editcuentabancaria':
            try:
                persona = request.session['persona']
                f = CuentaBancariaPersonaForm(request.POST)
                if f.is_valid():
                    cuentabancaria = CuentaBancariaPersona.objects.get(pk=int(encrypt(request.POST['id'])))
                    if persona.cuentabancariapersona_set.filter(numero=f.cleaned_data['numero'].strip()).exclude(
                            id=cuentabancaria.id).exists():
                        return JsonResponse({'result': True,
                                             "mensaje": "La cuenta bancaria se encuentra registrada."})

                    if cuentabancaria.verificado:
                        return JsonResponse({'result': True,
                                             "mensaje": "mensaje': u'No puede modificar la cuenta bancaria."})
                    cuentabancaria.numero = f.cleaned_data['numero']
                    cuentabancaria.banco = f.cleaned_data['banco']
                    cuentabancaria.tipocuentabanco = f.cleaned_data['tipocuentabanco']
                    cuentabancaria.save(request)
                    if 'archivo' in request.FILES:
                        newfile = request.FILES['archivo']
                        if newfile:
                            newfile._name = generar_nombre("cuentabancaria_", newfile._name)
                            cuentabancaria.archivo = newfile
                            cuentabancaria.save(request)
                    log(u'Modifico cuenta bancaria: %s' % persona, request, "edit")
                    return JsonResponse({'result': False, 'mensaje': 'Guardado con exito'})
                else:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()],
                                         "mensaje": "Error en el formulario"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': f'{ex}'})

        elif action == 'delcuentabancaria':
            try:
                cuentabancaria = CuentaBancariaPersona.objects.get(pk=int(encrypt(request.POST['id'])))
                if cuentabancaria.verificado:
                    res_js = {'error': True, 'mensaje': 'No puede eliminar la cuenta bancaria.'}
                    return JsonResponse(res_js)
                if cuentabancaria.activapago:
                    res_js = {'error': True, 'mensaje': 'No puede eliminar la cuenta bancaria se encuentra asignada al proceso de beca.'}
                    return JsonResponse(res_js)
                cuentabancaria.status = False
                cuentabancaria.save(request)
                log(u'Elimino cuenta bancaria: %s' % cuentabancaria, request, "del")
                res_js = {'error': False}
            except Exception as ex:
                transaction.set_rollback(True)
                err = 'Error on line {}'.format(sys.exc_info()[-1].tb_lineno)
                msg_err = f"Ocurrio un error: {ex.__str__()}. {err}"
                res_js = {'error': True, 'mensaje': msg_err}
            return JsonResponse(res_js)

        elif action == 'verificarcuenta':
            with transaction.atomic():
                try:
                    val = eval(request.POST['val'].capitalize())
                    registro = CuentaBancariaPersona.objects.get(pk=int(request.POST['id']))
                    registro.verificado = val
                    registro.save(request)
                    log(u'Verifico cuenta: %s - %s' % (registro,val), request, "edit")
                    return JsonResponse({"result": True, 'mensaje': 'Cambios guardados'})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": False})

        # Declaración de bienes

        elif action == 'adddeclaracion':
            try:
                newfile = None
                if 'archivo' in request.FILES:
                    newfile = request.FILES['archivo']
                    if newfile.size > 4194304:
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"Error, Tamaño de archivo Maximo permitido es de 4Mb"})
                administrativo = Persona.objects.get(id=request.POST['id'])
                f = DeclaracionBienForm(request.POST, request.FILES)
                if f.is_valid():
                    if 'archivo' not in request.FILES:
                        return JsonResponse({'result': True, 'mensaje': u'Debe subir la declaración en formato pdf'})
                    if DeclaracionBienes.objects.filter(persona=administrativo, fecha=f.cleaned_data['fecha'],status=True, tipodeclaracion=f.cleaned_data['tipodeclaracion']).exists():
                        return JsonResponse({'result': True, 'mensaje': u'Ya tiene una declaración en la misma fecha'})
                    newfile = request.FILES['archivo']
                    newfile._name = generar_nombre("declaracion_", newfile._name)
                    fechaproxima = None
                    if int(f.cleaned_data['tipodeclaracion']) == 3:
                        fechaproxima = f.cleaned_data['fecha'] + timedelta(days=730)  # sumar dos años

                    declaracion = DeclaracionBienes(persona=administrativo,
                                                    fecha=f.cleaned_data['fecha'],
                                                    tipodeclaracion=f.cleaned_data['tipodeclaracion'],
                                                    codigobarra=f.cleaned_data['codigobarra'],
                                                    fechaproximoregistro=fechaproxima,
                                                    archivo=newfile)
                    declaracion.save(request)
                    if newfile is not None:
                        if DistributivoPersona.objects.filter(status=True, persona=administrativo).exists():
                            lista = ['talento_humano@unemi.edu.ec', 'portizg@unemi.edu.ec']
                            asunto = "Ingresaron nueva declaración (archivo)"
                            send_html_mail(asunto, "emails/ingreso_archivos_tthh.html",
                                           {'asunto': asunto, 'd': declaracion.persona.nombre_completo_inverso(),
                                            'fecha': datetime.now().date(), 'escenario': 'declaración'}, lista, [],
                                           cuenta=CUENTAS_CORREOS[1][1])
                        log(u'Adiciono declaracion: %s' % administrativo, request, "add")
                    return JsonResponse({'result': False, 'mensaje': 'Guardado con exito'})
                else:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()],
                                             "mensaje": "Error en el formulario"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': f'{ex}'})

        elif action == 'editdeclaracion':
            try:
                newfile = None
                if 'archivo' in request.FILES:
                    newfile = request.FILES['archivo']
                    if newfile.size > 4194304:
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"Error, Tamaño de archivo Maximo permitido es de 4Mb"})
                declaracion = DeclaracionBienes.objects.get(id=request.POST['id'])
                f = DeclaracionBienForm(request.POST, request.FILES)
                if f.is_valid():
                    if 'archivo' in request.FILES:
                        newfile = request.FILES['archivo']
                        newfile._name = generar_nombre("declaracion_", newfile._name)
                        declaracion.archivo = newfile

                    fechaproxima = None
                    if int(f.cleaned_data['tipodeclaracion']) == 3:
                        fechaproxima = f.cleaned_data['fecha'] + timedelta(days=730)  # sumar dos años
                    declaracion.fecha=f.cleaned_data['fecha']
                    declaracion.tipodeclaracion=f.cleaned_data['tipodeclaracion']
                    declaracion.codigobarra=f.cleaned_data['codigobarra']
                    declaracion.fechaproximoregistro=fechaproxima
                    declaracion.save(request)
                    if newfile is not None:
                        if DistributivoPersona.objects.filter(status=True, persona=administrativo).exists():
                            lista = ['talento_humano@unemi.edu.ec', 'portizg@unemi.edu.ec']
                            asunto = "Ingresaron nueva declaración (archivo)"
                            send_html_mail(asunto, "emails/ingreso_archivos_tthh.html",
                                           {'asunto': asunto, 'd': declaracion.persona.nombre_completo_inverso(),
                                            'fecha': datetime.now().date(), 'escenario': 'declaración'}, lista, [],
                                           cuenta=CUENTAS_CORREOS[1][1])
                        log(u'Adiciono declaracion: %s' % administrativo, request, "add")
                    return JsonResponse({'result': False, 'mensaje': 'Guardado con exito'})
                else:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()],
                                             "mensaje": "Error en el formulario"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': f'{ex}'})

        elif action == 'deldeclaracion':
            try:
                declaracion = DeclaracionBienes.objects.get(pk=int(encrypt(request.POST['id'])))
                if declaracion.verificado:
                    res_js = {'error': True, 'mensaje': 'No puede eliminar la cuenta bancaria.'}
                    return JsonResponse(res_js)
                declaracion.status = False
                declaracion.save(request)
                log(u'Elimino declaracion de bienes: %s' % declaracion, request, "del")
                res_js = {'error': False}
            except Exception as ex:
                transaction.set_rollback(True)
                err = 'Error on line {}'.format(sys.exc_info()[-1].tb_lineno)
                msg_err = f"Ocurrio un error: {ex.__str__()}. {err}"
                res_js = {'error': True, 'mensaje': msg_err}
            return JsonResponse(res_js)

        elif action == 'verificardeclaracion':
            with transaction.atomic():
                try:
                    val = eval(request.POST['val'].capitalize())
                    registro = DeclaracionBienes.objects.get(pk=int(request.POST['id']))
                    registro.verificado = val
                    registro.save(request)
                    log(u'Verifico declaracion: %s - %s' % (registro,val), request, "edit")
                    return JsonResponse({"result": True, 'mensaje': 'Cambios guardados'})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": False})

        # Referencias

        elif action == 'addreferencia':
            try:
                form = ReferenciaPersonaForm(request.POST)
                administrativo = Persona.objects.get(id=request.POST['id'])
                if form.is_valid():
                    referenciapersona = ReferenciaPersona(persona=administrativo,
                                                          nombres=form.cleaned_data['nombres'],
                                                          apellidos=form.cleaned_data['apellidos'],
                                                          email=form.cleaned_data['email'],
                                                          telefono=form.cleaned_data['telefono'],
                                                          institucion=form.cleaned_data['institucion'],
                                                          relacion=form.cleaned_data['relacion'],
                                                          cargo=form.cleaned_data['cargo'])
                    referenciapersona.save(request)
                    log(u'Adicionó referencia personal a la hoja de vida: %s - la persona: %s' % (
                        referenciapersona, administrativo), request, "add")
                    return JsonResponse({'result': False, 'mensaje': 'Guardado con exito'})
                else:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()],
                                         "mensaje": "Error en el formulario"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', 'mensaje': f'{ex}'})

        elif action == 'editreferencia':
            try:
                form = ReferenciaPersonaForm(request.POST)
                if form.is_valid():
                    referenciapersona = ReferenciaPersona.objects.get(pk=(request.POST['id']))
                    referenciapersona.nombres = form.cleaned_data['nombres']
                    referenciapersona.apellidos = form.cleaned_data['apellidos']
                    referenciapersona.email = form.cleaned_data['email']
                    referenciapersona.telefono = form.cleaned_data['telefono']
                    referenciapersona.institucion = form.cleaned_data['institucion']
                    referenciapersona.relacion = form.cleaned_data['relacion']
                    referenciapersona.cargo = form.cleaned_data['cargo']
                    referenciapersona.save(request)
                    log(u'Editó una referencia personal a la hoja de vida: %s - la persona: %s' % (
                        referenciapersona, referenciapersona.persona), request, "edit")
                    return JsonResponse({'result': False, 'mensaje': 'Guardado con exito'})
                else:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()],
                                         "mensaje": "Error en el formulario"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', 'mensaje': f'{ex}'})

        elif action == 'delreferencia':
            try:
                referenciapersona = ReferenciaPersona.objects.get(pk=encrypt(request.POST['id']))
                log(u'Eliminó un referencia personal a la hoja de vida: %s - la persona: %s' % (
                    referenciapersona, persona), request, "add")
                referenciapersona.status = False
                referenciapersona.save(request)
                res_js = {'error': False}
            except Exception as ex:
                transaction.set_rollback(True)
                err = 'Error on line {}'.format(sys.exc_info()[-1].tb_lineno)
                msg_err = f"Ocurrio un error: {ex.__str__()}. {err}"
                res_js = {'error': True, 'mensaje': msg_err}
            return JsonResponse(res_js)

        # Gastos personales

        elif action == 'subirproyecciongastos':
            try:
                file=request.FILES['archivo']
                gasto=GastosPersonales.objects.get(id=request.POST['id'])
                name = file._name
                ext = name[name.rfind("."):]
                if not ext == '.pdf':
                    raise NameError('Solo se permite formato pdf')
                file._name = generar_nombre(f"proyeccion_{gasto.persona.usuario.username}_{gasto.id}", "firmado")+'.pdf'
                gasto.archivo=file
                gasto.save(request)
                log(u'Guardo archivo firmado: {}'.format(gasto), request, "add")
                return JsonResponse({"result": False, "mensaje": "Guardado con exito", }, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": u"Error al guardar. %s" % ex.__str__()}, safe=False)

    #Formación Académica

        #Titulación

        elif action == 'addtitulacion':
            try:
                administrativo = Persona.objects.get(id=request.POST['id'])
                f = TitulacionPersonaForm(request.POST, request.FILES)
                if 'archivo' in request.FILES:
                    arch = request.FILES['archivo']
                    exte = ext_archive(arch._name)
                    if arch.size > 4194304:
                        return JsonResponse({"result": "bad", "mensaje": u"Error, archivo mayor a 4 Mb."})
                    if not exte in  ['.pdf', '.png','.jpg', '.jpeg','.jpg']:
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"Archivo Título solo en formato .pdf, jpg, jpeg, png"})
                if 'registroarchivo' in request.FILES:
                    registroarchivo = request.FILES['registroarchivo']
                    exte1 = ext_archive(registroarchivo._name)
                    if registroarchivo.size > 4194304:
                        return JsonResponse({"result": "bad", "mensaje": u"Error, archivo mayor a 4 Mb."})
                    if not exte1 in  ['.pdf', '.png','.jpg', '.jpeg','.jpg']:
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"Archivo SENESCYT solo en formato .pdf, jpg, jpeg, png"})
                if not f.is_valid():
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()],
                                         "mensaje": "Error en el formulario"})
                titulacion = Titulacion(persona=administrativo,
                                        titulo=f.cleaned_data['titulo'],
                                        areatitulo=f.cleaned_data['areatitulo'],
                                        fechainicio=f.cleaned_data['fechainicio'],
                                        fechaobtencion=f.cleaned_data['fechaobtencion'],
                                        fechaegresado=f.cleaned_data['fechaegresado'],
                                        registro=f.cleaned_data['registro'],
                                        pais=f.cleaned_data['pais'],
                                        provincia=f.cleaned_data['provincia'],
                                        canton=f.cleaned_data['canton'],
                                        parroquia=f.cleaned_data['parroquia'],
                                        educacionsuperior=f.cleaned_data['educacionsuperior'],
                                        institucion=f.cleaned_data['institucion'],
                                        colegio=f.cleaned_data['colegio'],
                                        anios=f.cleaned_data['anios'],
                                        semestres=f.cleaned_data['semestres'],
                                        cursando=f.cleaned_data['cursando'],
                                        aplicobeca=f.cleaned_data['aplicobeca'],
                                        tipobeca=f.cleaned_data['tipobeca'] if f.cleaned_data[
                                            'aplicobeca'] else None,
                                        financiamientobeca=f.cleaned_data['financiamientobeca'] if f.cleaned_data[
                                            'aplicobeca'] else None,
                                        valorbeca=f.cleaned_data['valorbeca'] if f.cleaned_data[
                                            'aplicobeca'] else 0)
                titulacion.save(request)
                if not titulacion.cursando:
                    if 'archivo' in request.FILES:
                        newfile = request.FILES['archivo']
                        if newfile:
                            newfile._name = generar_nombre("titulacion_", newfile._name)
                            titulacion.archivo = newfile
                            titulacion.save(request)
                        if DistributivoPersona.objects.filter(status=True, persona=administrativo):
                            lista = ['talento_humano@unemi.edu.ec', 'portizg@unemi.edu.ec']
                            asunto = "Ingresaron título académico (archivo)"
                            send_html_mail(asunto, "emails/ingreso_archivos_tthh.html",
                                           {'asunto': asunto, 'd': titulacion.persona.nombre_completo_inverso(),
                                            'fecha': datetime.now().date(), 'escenario': 'titulación - SENECYT'},
                                           lista, [], cuenta=CUENTAS_CORREOS[1][1])

                    if 'registroarchivo' in request.FILES:
                        newfile2 = request.FILES['registroarchivo']
                        if newfile2:
                            newfile2._name = generar_nombre("archivosenecyt_", newfile2._name)
                            titulacion.registroarchivo = newfile2
                            titulacion.save(request)
                        if DistributivoPersona.objects.filter(status=True, persona=persona):
                            lista = ['talento_humano@unemi.edu.ec', 'portizg@unemi.edu.ec']
                            asunto = "Ingresaron título académico (archivo SENESCYT)"
                            send_html_mail(asunto, "emails/ingreso_archivos_tthh.html",
                                           {'asunto': asunto, 'd': titulacion.persona.nombre_completo_inverso(),
                                            'fecha': datetime.now().date(),
                                            'escenario': 'titulación - título academico'}, lista, [],
                                           cuenta=CUENTAS_CORREOS[1][1])
                # campotitulo = None
                # if CamposTitulosPostulacion.objects.filter(status=True, titulo=f.cleaned_data['titulo']).exists():
                #     campotitulo = CamposTitulosPostulacion.objects.filter(status=True,
                #                                                           titulo=f.cleaned_data['titulo']).first()
                # else:
                #     campotitulo = CamposTitulosPostulacion(titulo=f.cleaned_data['titulo'])
                #     campotitulo.save(request)
                # for ca in f.cleaned_data['campoamplio']:
                #     if not campotitulo.campoamplio.filter(id=ca.id):
                #         campotitulo.campoamplio.add(ca)
                # for ce in f.cleaned_data['campoespecifico']:
                #     if not campotitulo.campoespecifico.filter(id=ce.id):
                #         campotitulo.campoespecifico.add(ce)
                # for cd in f.cleaned_data['campodetallado']:
                #     if not campotitulo.campodetallado.filter(id=cd.id):
                #         campotitulo.campodetallado.add(cd)
                # campotitulo.save()
                log(u'Adiciono titulacion: %s' % persona, request, "add")
                return JsonResponse({'result': False, 'mensaje': 'Guardado con exito'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': f'Error al guardar los datos {ex}'})

        elif action == 'edittitulacion':
            try:
                administrativo = Persona.objects.get(id=request.POST['idp'])
                f = TitulacionPersonaForm(request.POST, request.FILES)
                if 'archivo' in request.FILES:
                    arch = request.FILES['archivo']
                    extencion = arch._name.split('.')
                    exte = extencion[1]
                    if arch.size > 4194304:
                        return JsonResponse({"result": "bad", "mensaje": u"Error, archivo mayor a 4 Mb."})
                    if not exte == 'pdf' and not exte == 'png' and not exte == 'jpg' and not exte == 'jpeg' and not exte == 'jpg':
                        return JsonResponse({"result": "bad", "mensaje": u"Solo archivos .pdf, jpg, jpeg, png"})

                if 'registroarchivo' in request.FILES:
                    registroarchivo = request.FILES['registroarchivo']
                    extencion1 = registroarchivo._name.split('.')
                    exte1 = extencion1[1]
                    if registroarchivo.size > 4194304:
                        return JsonResponse({"result": "bad", "mensaje": u"Error, archivo mayor a 4 Mb."})
                    if not exte1 == 'pdf' and not exte1 == 'png' and not exte1 == 'jpg' and not exte1 == 'jpeg' and not exte1 == 'jpg':
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"Archivo SENESCYT solo en formato .pdf, jpg, jpeg, png"})
                if not f.is_valid():
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()],
                                         "mensaje": "Error en el formulario"})
                titulacion = Titulacion.objects.get(pk=encrypt_id(request.POST['id']))
                titulacion.areatitulo = f.cleaned_data['areatitulo']
                titulacion.fechainicio = f.cleaned_data['fechainicio']
                titulacion.fechaobtencion = f.cleaned_data['fechaobtencion']
                titulacion.fechaegresado = f.cleaned_data['fechaegresado']
                titulacion.pais = f.cleaned_data['pais']
                titulacion.provincia = f.cleaned_data['provincia']
                titulacion.canton = f.cleaned_data['canton']
                titulacion.parroquia = f.cleaned_data['parroquia']
                titulacion.colegio = f.cleaned_data['colegio']
                titulacion.anios = f.cleaned_data['anios']
                titulacion.semestres = f.cleaned_data['semestres']
                titulacion.aplicobeca = f.cleaned_data['aplicobeca']
                titulacion.tipobeca = f.cleaned_data['tipobeca'] if f.cleaned_data['aplicobeca'] else None
                titulacion.financiamientobeca = f.cleaned_data['financiamientobeca'] if f.cleaned_data[
                    'aplicobeca'] else None
                titulacion.valorbeca = f.cleaned_data['valorbeca'] if f.cleaned_data['aplicobeca'] else 0
                if not titulacion.verificadosenescyt:
                    titulacion.titulo = f.cleaned_data['titulo']
                    titulacion.educacionsuperior = f.cleaned_data['educacionsuperior']
                    titulacion.institucion = f.cleaned_data['institucion']
                    titulacion.registro = f.cleaned_data['registro']
                    titulacion.cursando = f.cleaned_data['cursando']
                if 'archivo' in request.FILES:
                    newfile = request.FILES['archivo']
                    newfile._name = generar_nombre("titulacion_", newfile._name)
                    titulacion.archivo = newfile
                if 'registroarchivo' in request.FILES:
                    newfile2 = request.FILES['registroarchivo']
                    if newfile2:
                        newfile2._name = generar_nombre("archivosenecyt_", newfile2._name)
                        titulacion.registroarchivo = newfile2
                titulacion.save(request)
                request.session['instruccion'] = 1
                if Graduado.objects.filter(status=True, inscripcion__persona__id=administrativo.id).exists():
                    datos = Persona.objects.get(status=True, id=persona.id)
                    if 'personales' in request.session and 'nacimiento' in request.session and 'domicilio' in request.session and 'etnia' in request.session and 'instruccion' in request.session and datos.datosactualizados == 0:
                        if request.session['personales'] == 1 and request.session['nacimiento'] == 1 and \
                                request.session['domicilio'] == 1 and request.session['etnia'] == 1 and \
                                request.session['instruccion'] == 1 and datos.datosactualizados == 0:
                            datos.datosactualizados = 1
                            datos.save(request)
                log(u'Modifico titulacion: %s' % persona, request, "edit")
                return JsonResponse({'result': False, 'mensaje': 'Guardado con exito'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': f'Error al guardar los datos {ex}'})

        elif action == 'deltitulacion':
            try:
                persona = request.session['persona']
                titulacion = Titulacion.objects.get(pk=encrypt_id(request.POST['id']))
                if titulacion.verificado:
                    return JsonResponse({'result': 'bad', 'mensaje': u'No puede eliminar el titulo.'})
                log(u'Elimino titulacion: %s' % titulacion, request, "del")
                titulacion.status = False
                titulacion.save()
                res_js = {'error': False}
            except Exception as ex:
                transaction.set_rollback(True)
                err = 'Error on line {}'.format(sys.exc_info()[-1].tb_lineno)
                msg_err = f"Ocurrio un error: {ex.__str__()}. {err}"
                res_js = {'error': True, 'mensaje': msg_err}
            return JsonResponse(res_js)

        elif action == 'tituloprincipal':
            try:
                titulacion = Titulacion.objects.get(pk=encrypt_id(request.POST['id']))
                val=eval(request.POST['val'].capitalize())
                if val:
                    # if not titulacion.verificado:
                    #     return JsonResponse({'result': 'bad', 'mensaje': u'No se ha verificado su titulo'})
                    titulacion.persona.titulacion_set.update(principal=False)
                    titulacion.principal = val
                else:
                    titulacion.principal = val
                titulacion.save(request)
                log(u"Selecciono titulo principal: %s" % titulacion, request, "del")
                return JsonResponse({'result': True, 'mensaje': 'Guardado con exito'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': False, 'mensaje': f'{ex}'})

        elif action == 'adddetalletitulobachiller':
            try:
                f = DetalleTitulacionBachillerForm(request.POST, request.FILES)
                administrativo = Persona.objects.get(id=request.POST['id'])

                if not f.is_valid():
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()],
                                         "mensaje": "Error en el formulario"})
                titulacion = Titulacion.objects.get(pk=encrypt_id(request.POST['id']))
                newfile = None
                newfile2 = None
                if 'archivo' in request.FILES:
                    newfile = request.FILES['archivo']
                    if newfile:
                        newfile._name = generar_nombre("actagradobachiller_", newfile._name)
                # else:
                #     h = Truncator(request.POST['archivo1']).chars(-30, truncate = '/')
                #     newfile = f.cleaned_data['archivo1']

                if 'reconocimientoacademico' in request.FILES:
                    newfile2 = request.FILES['reconocimientoacademico']

                    if newfile2:
                        newfile2._name = generar_nombre("reconocimientoacademico_", newfile2._name)
                # else:
                #     newfile2 =f.cleaned_data['reconocimientoacademico1']

                if DetalleTitulacionBachiller.objects.filter(titulacion=titulacion).exists():
                    detalle = DetalleTitulacionBachiller.objects.filter(titulacion=titulacion)[0]
                    detalle.titulacion = titulacion
                    detalle.calificacion = f.cleaned_data['calificacion']
                    detalle.actagrado = newfile if newfile else detalle.actagrado
                    detalle.anioinicioperiodograduacion = f.cleaned_data['anioinicioperiodograduacion']
                    detalle.aniofinperiodograduacion = f.cleaned_data['aniofinperiodograduacion']
                    detalle.reconocimientoacademico = newfile2 if newfile2 else detalle.reconocimientoacademico
                    detalle.save(request)
                    titulacion.registroarchivo = newfile if newfile else detalle.actagrado
                    titulacion.save(request)

                else:
                    detalle = DetalleTitulacionBachiller(titulacion=titulacion,
                                                         calificacion=f.cleaned_data['calificacion'],
                                                         actagrado=newfile,
                                                         anioinicioperiodograduacion=f.cleaned_data[
                                                             'anioinicioperiodograduacion'],
                                                         aniofinperiodograduacion=f.cleaned_data[
                                                             'aniofinperiodograduacion'],
                                                         reconocimientoacademico=newfile2)
                    titulacion.registroarchivo = newfile
                    titulacion.save(request)

                    detalle.save(request)

                # para tambien porder registrar en archivo
                archivobachiller = Archivo.objects.filter(tipo_id=16, inscripcion__persona=persona, status=True)
                archivoreconocimiento = Archivo.objects.filter(tipo_id=18, inscripcion__persona=persona,
                                                               status=True)
                inscri = administrativo.perfilprincipal.inscripcion if administrativo.perfilprincipal.es_estudiante() else None
                profe = administrativo.perfilprincipal.inscripcion if administrativo.perfilprincipal.es_profesor() else None

                if not archivobachiller and newfile:  # si no exsite en Archivo, lo guarda en tabla archivo sino lo actualiza
                    nombrearchivo = f'ACTA DE GRADO DE BACHILLER DE LA PERSONA: {persona}'

                    archivobachi = Archivo(
                        tipo_id=16,
                        nombre=nombrearchivo,
                        fecha=datetime.now().date(),
                        archivo=newfile,
                        aprobado=True,
                        inscripcion=inscri,
                        profesor=profe,
                        persona=persona,
                        sga=True
                    )
                    archivobachi.save(request)
                else:
                    if newfile:
                        archivobachi = archivobachiller.last()
                        archivobachi.inscripcion = inscri
                        archivobachi.archivo = newfile
                        archivobachi.save(request)

                if not archivoreconocimiento and newfile2:  # si no exsite en Archivo, lo guarda en tabla archivo sino lo actualiza
                    nombrearchivo = f'RECONOCIMIENTO ACADÉMICO DE LA PERSONA: {persona}'

                    archivorecono = Archivo(
                        tipo_id=18,
                        nombre=nombrearchivo,
                        fecha=datetime.now().date(),
                        archivo=newfile2,
                        aprobado=True,
                        inscripcion=inscri,
                        profesor=profe,
                        persona=persona,
                        sga=True
                    )
                    archivorecono.save(request)
                else:
                    if newfile2:
                        archivorecono = archivoreconocimiento.last()
                        archivorecono.inscripcion = inscri
                        archivorecono.archivo = newfile2
                        archivorecono.save(request)

                messages.success(request, 'Se guardó exitosamente.')
                log(u'Adiciono titulacion Bachiller: %s %s %s %s' % (
                    detalle, detalle.calificacion, detalle.anioinicioperiodograduacion,
                    detalle.aniofinperiodograduacion), request, "add")
                return JsonResponse({'result': False, 'mensaje': 'Guardado con exito'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': f'Error al guardar los datos {ex}'})

        elif action == 'verificartitulacion':
            with transaction.atomic():
                try:
                    val = eval(request.POST['val'].capitalize())
                    registro = Titulacion.objects.get(pk=int(request.POST['id']))
                    registro.verificado = val
                    if val:
                        registro.personaaprobaciontitulo=persona
                        registro.fechaaprobaciontitulo=datetime.now()
                    else:
                        registro.personaaprobaciontitulo = None
                        registro.fechaaprobaciontitulo = None
                    registro.save(request)
                    log(u'Verificó (%s) titulación : %s (%s)- persona: %s ' % (val,registro, registro.id, registro.persona), request, "edit")
                    return JsonResponse({"result": True, 'mensaje': 'Cambios guardados'})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": False})

        elif action == 'tituloverisenescyt':
            try:
                val = eval(request.POST['val'].capitalize())
                titulacion = Titulacion.objects.get(pk=int(request.POST['id']))
                titulacion.verisenescyt = val
                log(u"Marca (%s) titulo registrado de la senescyt: %s (%s) - persona: %s" % (val,
                titulacion, titulacion.id, titulacion.persona), request, "edit")
                titulacion.save(request)
                return JsonResponse({"result": True, 'mensaje': 'Cambios guardados'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": False})

        elif action == 'tituloveriarchivotitulo':
            try:
                val = eval(request.POST['val'].capitalize())
                titulacion = Titulacion.objects.get(pk=int(request.POST['id']))
                titulacion.veriarchivotitulo = val
                log(u"Selecciona (%s) archivo titulo : %s (%s) - persona: %s" % (val,
                titulacion, titulacion.id, titulacion.persona), request, "edit")
                titulacion.save(request)
                return JsonResponse({"result": True, 'mensaje': 'Cambios guardados'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": False})


        #Capacitación

        elif action == 'addcapacitacion':
            try:
                administrativo = Persona.objects.get(id=request.POST['id'])
                f = CapacitacionPersonaForm(request.POST, request.FILES)
                if 'archivo' in request.FILES:
                    d = request.FILES['archivo']
                    if d.size > 4194304:
                        raise NameError("Error, archivo mayor a 4 Mb.")
                    else:
                        newfiles = request.FILES['archivo']
                        newfilesd = newfiles._name
                        ext = newfilesd[newfilesd.rfind("."):]
                        if ext == '.pdf' or ext == '.PDF' or ext == '.png' or ext == '.jpg' or ext == '.jpeg':
                            tipocapacitacion = int(request.POST['tipocapacitacion'])
                            if tipocapacitacion == 2:
                                f.fields['anio'].required = False
                                anio = datetime.now().date().year
                            if not f.is_valid():
                                transaction.set_rollback(True)
                                return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()],
                                                     "mensaje": "Error en el formulario"})
                            newfile = request.FILES['archivo']
                            newfile._name = generar_nombre("capacitacion_", newfile._name)
                            anio = f.cleaned_data['anio']
                            capacitacion = Capacitacion(persona=administrativo,
                                                        institucion=f.cleaned_data['institucion'],
                                                        nombre=f.cleaned_data['nombre'],
                                                        descripcion=f.cleaned_data['descripcion'],
                                                        tipo=f.cleaned_data['tipo'],
                                                        tipocurso=f.cleaned_data['tipocurso'],
                                                        tipocapacitacion=f.cleaned_data['tipocapacitacion'],
                                                        tipocertificacion=f.cleaned_data['tipocertificacion'],
                                                        tipoparticipacion=f.cleaned_data['tipoparticipacion'],
                                                        anio=anio,
                                                        contextocapacitacion=f.cleaned_data['contexto'],
                                                        detallecontextocapacitacion=f.cleaned_data['detallecontexto'],
                                                        auspiciante=f.cleaned_data['auspiciante'],
                                                        areaconocimiento=f.cleaned_data['areaconocimiento'],
                                                        subareaconocimiento=f.cleaned_data['subareaconocimiento'],
                                                        subareaespecificaconocimiento=f.cleaned_data['subareaespecificaconocimiento'],
                                                        pais=f.cleaned_data['pais'],
                                                        provincia=f.cleaned_data['provincia'],
                                                        canton=f.cleaned_data['canton'],
                                                        parroquia=f.cleaned_data['parroquia'],
                                                        fechainicio=f.cleaned_data['fechainicio'],
                                                        fechafin=f.cleaned_data['fechafin'],
                                                        horas=f.cleaned_data['horas'],
                                                        expositor=f.cleaned_data['expositor'],
                                                        modalidad=f.cleaned_data['modalidad'],
                                                        otramodalidad=f.cleaned_data['otramodalidad'],
                                                        archivo=newfile)
                            capacitacion.save(request)
                            if DistributivoPersona.objects.filter(status=True, persona=administrativo):
                                lista = ['talento_humano@unemi.edu.ec', 'portizg@unemi.edu.ec']
                                asunto = "Ingresaron nueva capacitación (archivo)"
                                send_html_mail(asunto, "emails/ingreso_archivos_tthh.html",
                                               {'asunto': asunto,
                                                'd': capacitacion.persona.nombre_completo_inverso(),
                                                'fecha': datetime.now().date(), 'escenario': 'capacitación'}, lista,
                                               [],
                                               cuenta=CUENTAS_CORREOS[1][1])

                            log(u'Adiciono capacitacion: %s' % persona, request, "add")
                            return JsonResponse({'result': False, 'mensaje': 'Guardado con exito'})
                        else:
                            raise NameError("Error, solo archivos .pdf, png, jpg, jpeg.")
                else:
                    raise NameError("No existe archivo")
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': f'{ex}'})

        elif action == 'editcapacitacion':
            try:
                persona = request.session['persona']
                capacitacion = Capacitacion.objects.get(pk=encrypt_id(request.POST['id']))
                if not capacitacion.archivo and not 'archivo' in request.FILES:
                    raise NameError('No ha seleccionado archivo')
                if 'archivo' in request.FILES:
                    d = request.FILES['archivo']
                    if d.size > 4194304:
                        raise NameError("Error, archivo mayor a 4 Mb.")
                    else:
                        newfiles = request.FILES['archivo']
                        newfilesd = newfiles._name
                        ext = newfilesd[newfilesd.rfind("."):]
                        if not (ext == '.pdf' or ext == '.png' or ext == '.jpg' or ext == '.jpeg'):
                            raise NameError("Error, solo archivos .pdf, png, jpg, jpeg.")
                f = CapacitacionPersonaForm(request.POST, request.FILES, instancia=capacitacion)
                tipocapacitacion = int(request.POST['tipocapacitacion'])
                if tipocapacitacion == 2:
                    f.fields['anio'].required = False
                    anio = datetime.now().date().year
                if not f.is_valid():
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()],
                                         "mensaje": "Error en el formulario"})
                anio = f.cleaned_data['anio']
                capacitacion.institucion = f.cleaned_data['institucion']
                capacitacion.nombre = f.cleaned_data['nombre']
                capacitacion.descripcion = f.cleaned_data['descripcion']
                capacitacion.tipo = f.cleaned_data['tipo']
                capacitacion.tipocurso = f.cleaned_data['tipocurso']
                capacitacion.tipocapacitacion = f.cleaned_data['tipocapacitacion']
                capacitacion.tipocertificacion = f.cleaned_data['tipocertificacion']
                capacitacion.tipoparticipacion = f.cleaned_data['tipoparticipacion']
                capacitacion.auspiciante = f.cleaned_data['auspiciante']
                capacitacion.areaconocimiento = f.cleaned_data['areaconocimiento']
                capacitacion.subareaconocimiento = f.cleaned_data['subareaconocimiento']
                capacitacion.subareaespecificaconocimiento = f.cleaned_data['subareaespecificaconocimiento']
                capacitacion.pais = f.cleaned_data['pais']
                capacitacion.anio = anio
                capacitacion.contextocapacitacion = f.cleaned_data['contexto']
                capacitacion.detallecontextocapacitacion = f.cleaned_data['detallecontexto']
                capacitacion.provincia = f.cleaned_data['provincia']
                capacitacion.canton = f.cleaned_data['canton']
                capacitacion.parroquia = f.cleaned_data['parroquia']
                capacitacion.fechainicio = f.cleaned_data['fechainicio']
                capacitacion.fechafin = f.cleaned_data['fechafin']
                capacitacion.horas = f.cleaned_data['horas']
                capacitacion.expositor = f.cleaned_data['expositor']
                capacitacion.modalidad = f.cleaned_data['modalidad']
                capacitacion.otramodalidad = f.cleaned_data['otramodalidad']
                if 'archivo' in request.FILES:
                    newfile = request.FILES['archivo']
                    newfile._name = generar_nombre("capacitacion_", newfile._name)
                    capacitacion.archivo = newfile
                # capacitacion.tiempo = f.cleaned_data['tiempo']
                capacitacion.save(request)
                log(u'Modifoco capacitacion: %s' % persona, request, "edit")
                return JsonResponse({'result': False, 'mensaje':'Guardado con éxito'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': f'{ex}'})

        elif action == 'delcapacitacion':
            try:
                capacitacion = Capacitacion.objects.get(pk=encrypt_id(request.POST['id']))
                if capacitacion.verificado:
                    raise NameError('No puede eliminar la capacitacion.')
                log(u'Elimino capacitacion: %s' % capacitacion, request, "del")
                capacitacion.status=False
                capacitacion.save(request)
                res_js = {'error': False}
            except Exception as ex:
                transaction.set_rollback(True)
                err = 'Error on line {}'.format(sys.exc_info()[-1].tb_lineno)
                msg_err = f"Ocurrio un error: {ex.__str__()}. {err}"
                res_js = {'error': True, 'mensaje': msg_err}
            return JsonResponse(res_js)

        elif action == 'deleventoaprobacion':
            try:
                cabecera = CapCabeceraSolicitud.objects.get(pk=encrypt_id(request.POST['id']))
                cabecera.capdetallesolicitud_set.filter(status=True).update(status=False)
                cabecera.status=False
                cabecera.save(request)
                log(u'Elimino Cabecera Solicitud de Evento: %s' % cabecera, request, "del")
                res_js = {'error': False}
            except Exception as ex:
                transaction.set_rollback(True)
                err = 'Error on line {}'.format(sys.exc_info()[-1].tb_lineno)
                msg_err = f"Ocurrio un error: {ex.__str__()}. {err}"
                res_js = {'error': True, 'mensaje': msg_err}
            return JsonResponse(res_js)

        elif action == 'verificarcapacitacion':
            with transaction.atomic():
                try:
                    val = eval(request.POST['val'].capitalize())
                    registro = Capacitacion.objects.get(pk=int(request.POST['id']))
                    registro.verificado = val
                    if val:
                        registro.verificado = True
                        registro.personaaprobacioncapacitacion = persona
                        registro.fechaaprobacioncapacitacion = datetime.now()
                    else:
                        registro.personaaprobacioncapacitacion = None
                        registro.fechaaprobacioncapacitacion = None
                    registro.save(request)
                    log(u'Verificó (%s) capacitacion : %s (%s)- persona: %s ' % (val,registro, registro.id, registro.persona), request, "edit")
                    registro.mail_notificar_aprobacion_uath(request.session['nombresistema'], persona)

                    return JsonResponse({"result": True, 'mensaje': 'Cambios guardados'})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": False})

        # Certificación

        elif action == 'addcertificadoidioma':
            try:
                administrativo = Persona.objects.get(id=request.POST['id'])
                f = CertificadoIdiomaForm(request.POST, request.FILES)
                if 'archivo' in request.FILES:
                    d = request.FILES['archivo']
                    newfilesd = d._name
                    ext = newfilesd[newfilesd.rfind("."):]
                    if ext == '.pdf' or ext == '.PDF' or ext == '.png' or ext == '.jpg' or ext == '.jpeg':
                        a = 1
                    else:
                        return JsonResponse({"result": "bad", "mensaje": u"Error, solo archivos .pdf."})
                    if d.size > 2194304:
                        return JsonResponse({"result": "bad", "mensaje": u"Error, archivo mayor a 2 Mb."})
                if not f.is_valid():
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()],
                                         "mensaje": "Error en el formulario"})
                certificado = CertificadoIdioma(persona=administrativo,
                                                idioma=f.cleaned_data['idioma'],
                                                institucioncerti=f.cleaned_data['institucioncerti'],
                                                validainst=f.cleaned_data['validainst'],
                                                otrainstitucion=f.cleaned_data['otrainstitucion'],
                                                nivelsuficencia=f.cleaned_data['nivelsuficencia'],
                                                fechacerti=f.cleaned_data['fechacerti'])
                certificado.save(request)
                newfile = None
                if 'archivo' in request.FILES:
                    newfile = request.FILES['archivo']
                    newfile._name = generar_nombre("certificado_otro_", newfile._name)
                    certificado.archivo = newfile
                    certificado.save(request)
                log(u'Adiciono certificado internacional: %s' % administrativo, request, "add")
                return JsonResponse({'result': False, 'mensaje': 'Guardado con exito'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': f'Error al guardar los datos: {ex}'})

        elif action == 'editcertificadoidioma':
            try:
                persona = request.session['persona']
                certificado = CertificadoIdioma.objects.get(pk=encrypt_id(request.POST['id']))
                if 'archivo' in request.FILES:
                    d = request.FILES['archivo']
                    if d.size > 2194304:
                        return JsonResponse({"result": "bad", "mensaje": u"Error, archivo mayor a 2 Mb."})
                    else:
                        newfiles = request.FILES['archivo']
                        newfilesd = newfiles._name
                        ext = newfilesd[newfilesd.rfind("."):]
                        if not (ext == '.pdf'):
                            return JsonResponse({"result": "bad", "mensaje": u"Error, solo archivos .pdf"})
                f = CertificadoIdiomaForm(request.POST, request.FILES)
                if not f.is_valid():
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()],
                                         "mensaje": "Error en el formulario"})
                certificado.idioma = f.cleaned_data['idioma']
                certificado.institucioncerti = f.cleaned_data['institucioncerti']
                certificado.otrainstitucion = f.cleaned_data['otrainstitucion']
                certificado.nivelsuficencia = f.cleaned_data['nivelsuficencia']
                certificado.fechacerti = f.cleaned_data['fechacerti']
                certificado.validainst = f.cleaned_data['validainst']
                if 'archivo' in request.FILES:
                    newfile = request.FILES['archivo']
                    newfile._name = generar_nombre("certificado_otro_", newfile._name)
                    certificado.archivo = newfile
                certificado.save(request)
                log(u'Modifico certificación de idioma personal: %s' % persona, request, "edit")
                return JsonResponse({'result': False, 'mensaje': 'Guardado con exito'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': f'Error al guardar los datos: {ex}'})

        elif action == 'delcertificadoidioma':
            try:
                certificadopersona = CertificadoIdioma.objects.get(pk=encrypt_id(request.POST['id']))
                certificadopersona.status = False
                certificadopersona.save(request)
                log(u'Eliminó un certificado internacional personal a la hoja de vida: %s - la persona: %s' % (certificadopersona, persona), request, "add")
                res_js = {'error': False}
            except Exception as ex:
                transaction.set_rollback(True)
                err = 'Error on line {}'.format(sys.exc_info()[-1].tb_lineno)
                msg_err = f"Ocurrio un error: {ex.__str__()}. {err}"
                res_js = {'error': True, 'mensaje': msg_err}
            return JsonResponse(res_js)

        elif action == 'addcertificado':
            try:
                administrativo = Persona.objects.get(id=request.POST['id'])
                f = CertificadoPersonaForm(request.POST, request.FILES)
                if 'archivo' in request.FILES:
                    d = request.FILES['archivo']
                    newfilesd = d._name
                    ext = newfilesd[newfilesd.rfind("."):]
                    if ext == '.pdf':
                        a = 1
                    else:
                        return JsonResponse({"result": "bad", "mensaje": u"Error, solo archivos .pdf."})
                    if d.size > 2194304:
                        return JsonResponse({"result": "bad", "mensaje": u"Error, archivo mayor a 2 Mb."})
                if not f.is_valid():
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()],
                                         "mensaje": "Error en el formulario"})
                certificado = CertificacionPersona(persona=administrativo,
                                                   nombres=f.cleaned_data['nombres'],
                                                   autoridad_emisora=f.cleaned_data['autoridad_emisora'],
                                                   numerolicencia=f.cleaned_data['numerolicencia'],
                                                   enlace=f.cleaned_data['enlace'],
                                                   vigente=f.cleaned_data['vigente'],
                                                   fechadesde=f.cleaned_data['fechadesde'],
                                                   fechahasta=f.cleaned_data['fechahasta'])
                certificado.save(request)
                newfile = None
                if 'archivo' in request.FILES:
                    newfile = request.FILES['archivo']
                    newfile._name = generar_nombre("certificado_", newfile._name)
                    certificado.archivo = newfile
                    certificado.save(request)
                log(u'Adiciono certificado: %s' % administrativo, request, "add")
                return JsonResponse({'result': False, 'mensaje': 'Guardado con exito'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': f'Error al guardar los datos {ex}'})

        elif action == 'editcertificado':
            try:
                persona = request.session['persona']
                certificado = CertificacionPersona.objects.get(pk=encrypt_id(request.POST['id']))
                if 'archivo' in request.FILES:
                    d = request.FILES['archivo']
                    if d.size > 2194304:
                        return JsonResponse({"result": "bad", "mensaje": u"Error, archivo mayor a 2 Mb."})
                    else:
                        newfiles = request.FILES['archivo']
                        newfilesd = newfiles._name
                        ext = newfilesd[newfilesd.rfind("."):]
                        if not (ext == '.pdf'):
                            return JsonResponse({"result": "bad", "mensaje": u"Error, solo archivos .pdf"})
                f = CertificadoPersonaForm(request.POST, request.FILES)
                if not f.is_valid():
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()],
                                         "mensaje": "Error en el formulario"})
                certificado.nombres = f.cleaned_data['nombres']
                certificado.autoridad_emisora = f.cleaned_data['autoridad_emisora']
                certificado.numerolicencia = f.cleaned_data['numerolicencia']
                certificado.fechadesde = f.cleaned_data['fechadesde']
                certificado.fechahasta = f.cleaned_data['fechahasta']
                certificado.enlace = f.cleaned_data['enlace']
                certificado.vigente = f.cleaned_data['vigente']
                if 'archivo' in request.FILES:
                    newfile = request.FILES['archivo']
                    newfile._name = generar_nombre("capacitacion_", newfile._name)
                    certificado.archivo = newfile
                # capacitacion.tiempo = f.cleaned_data['tiempo']
                certificado.save(request)
                log(u'Modifico certificación personal: %s' % persona, request, "edit")
                return JsonResponse({'result': False, 'mensaje': 'Guardado con exito'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': f'Error al guardar los datos {ex}'})

        elif action == 'delcertificado':
            try:
                certificadopersona = CertificacionPersona.objects.get(pk=encrypt_id(request.POST['id']))
                certificadopersona.status = False
                certificadopersona.save(request)
                log(u'Eliminó un certificado personal a la hoja de vida: %s - la persona: %s' % (certificadopersona, persona), request, "add")
                res_js = {'error': False}
            except Exception as ex:
                transaction.set_rollback(True)
                err = 'Error on line {}'.format(sys.exc_info()[-1].tb_lineno)
                msg_err = f"Ocurrio un error: {ex.__str__()}. {err}"
                res_js = {'error': True, 'mensaje': msg_err}
            return JsonResponse(res_js)

        elif action == 'verificarcertificacion':
            with transaction.atomic():
                try:
                    val = eval(request.POST['val'].capitalize())
                    registro = CertificacionPersona.objects.get(pk=int(request.POST['id']))
                    registro.verificado = val
                    if val:
                        registro.verificado = True
                        registro.personaaprobacioncertificado = persona
                        registro.fechaaprobacioncertificado = datetime.now()
                    else:
                        registro.personaaprobacioncertificado = None
                        registro.fechaaprobacioncertificado = None
                    registro.save(request)
                    log(u'Verificó (%s) certificacion : %s (%s)- persona: %s ' % (val,registro, registro.id, registro.persona), request, "edit")
                    registro.mail_notificar_aprobacion_uath(request.session['nombresistema'], persona)

                    return JsonResponse({"result": True, 'mensaje': 'Cambios guardados'})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": False})

        # Discapacidad

        elif action == 'editdiscapacidad':
            try:
                if 'archivo' in request.FILES:
                    arch = request.FILES['archivo']
                    extension = arch._name.split('.')
                    tam = len(extension)
                    exte = extension[tam - 1]
                    if arch.size > 4194304:
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"Error, el tamaño del archivo es mayor a 4 Mb."})
                    if not exte.lower() in ['pdf']:
                        return JsonResponse({"result": "bad", "mensaje": u"Solo se permiten archivos .pdf"})
                if 'archivovaloracion' in request.FILES:
                    arch = request.FILES['archivovaloracion']
                    extension = arch._name.split('.')
                    tam = len(extension)
                    exte = extension[tam - 1]
                    if arch.size > 4194304:
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"Error, el tamaño del archivo es mayor a 4 Mb."})
                    if not exte.lower() in ['pdf']:
                        return JsonResponse({"result": "bad", "mensaje": u"Solo se permiten archivos .pdf"})

                administrativo = Persona.objects.get(id=request.POST['id'])
                f = DiscapacidadForm(request.POST, request.FILES)
                if not f.is_valid():
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()],
                                         "mensaje": "Error en el formulario"})
                newfile = None
                if not f.cleaned_data['tienediscapacidad'] and f.cleaned_data['tienediscapacidadmultiple']:
                    raise NameError('No puede marcar discapacidad multiple sin marcar que tiene discapacidad')
                if not f.cleaned_data['tipodiscapacidad'] and f.cleaned_data['tipodiscapacidadmultiple']:
                    raise NameError(
                        'No puede elegir discapacidades multiples sin elegir una discapacidad principal')
                if f.cleaned_data['tienediscapacidadmultiple'] and not f.cleaned_data['tipodiscapacidadmultiple']:
                    raise NameError('Debe elegir una o más discapacidades multiples')
                perfil = administrativo.mi_perfil()
                perfil.tienediscapacidad = f.cleaned_data['tienediscapacidad']
                perfil.tipodiscapacidad = f.cleaned_data['tipodiscapacidad']
                perfil.porcientodiscapacidad = f.cleaned_data['porcientodiscapacidad'] if f.cleaned_data[
                    'porcientodiscapacidad'] else 0
                perfil.carnetdiscapacidad = f.cleaned_data['carnetdiscapacidad']
                perfil.institucionvalida = f.cleaned_data['institucionvalida']
                perfil.tienediscapacidadmultiple = f.cleaned_data['tienediscapacidadmultiple']
                perfil.grado = f.cleaned_data['grado'] if f.cleaned_data['grado'] else 0

                if not f.cleaned_data['tienediscapacidad']:
                    perfil.archivo = None
                    perfil.estadoarchivodiscapacidad = None
                    perfil.archivovaloracion = None
                if 'archivo' in request.FILES:
                    newfile = request.FILES['archivo']
                    newfile._name = generar_nombre("archivosdiscapacidad_", newfile._name)
                    perfil.archivo = newfile
                    perfil.estadoarchivodiscapacidad = 1
                if 'archivovaloracion' in request.FILES:
                    newfile = request.FILES['archivovaloracion']
                    newfile._name = generar_nombre("archivovaloracionmedica_", newfile._name)
                    perfil.archivovaloracion = newfile
                perfil.save(request)
                perfil.tipodiscapacidadmultiple.clear()
                perfil.subtipodiscapacidad.clear()
                if f.cleaned_data['tienediscapacidadmultiple']:
                    tipos = request.POST.getlist('tipodiscapacidadmultiple')
                    for tipo in tipos:
                        perfil.tipodiscapacidadmultiple.add(tipo)
                if f.cleaned_data['tipodiscapacidad'] and f.cleaned_data['tienediscapacidad']:
                    subtipos = request.POST.getlist('subtipodiscapacidad')
                    for subtipo in subtipos:
                        perfil.subtipodiscapacidad.add(subtipo)
                log(u'Modifico tipo de discapacidad: %s' % administrativo, request, "edit")
                return JsonResponse({'result': False, 'mensaje': 'Guardado con exito'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': f'Error al guardar los datos {ex}'})

        #Datos médicos


        elif action == 'editdatosmedicos':
            try:
                administrativo = Persona.objects.get(id=request.POST['id'])
                f = DatosMedicosForm(request.POST)
                if f.is_valid():
                    datosextension = administrativo.datos_extension()
                    examenfisico = administrativo.datos_examen_fisico()
                    datosextension.carnetiess = f.cleaned_data['carnetiess']
                    personaexamenfisico = \
                        PersonaExamenFisico.objects.filter(personafichamedica__personaextension=datosextension)[0]
                    personaexamenfisico.peso = f.cleaned_data['peso']
                    personaexamenfisico.talla = f.cleaned_data['talla']
                    personaexamenfisico.save(request)
                    persona.sangre = f.cleaned_data['sangre']
                    persona.save(request)
                    datosextension.save(request)

                    if 'archivotiposangre' in request.FILES:
                        newfile = request.FILES['archivotiposangre']
                        newfile._name = generar_nombre("tiposangre", newfile._name)

                        documento = administrativo.documentos_personales()
                        if documento is None:
                            documento = PersonaDocumentoPersonal(persona=administrativo,
                                                                 tiposangre=newfile,
                                                                 estadotiposangre=1
                                                                 )
                        else:
                            documento.tiposangre = newfile
                            documento.estadotiposangre = 1

                        documento.save(request)

                    log(u'Modifico datos de medicos basicos: %s' % administrativo, request, "edit")
                    return JsonResponse({'result': False, 'mensaje': 'Guardado con exito'})
                else:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()],
                                     "mensaje": "Error en el formulario"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', 'mensaje': f'{ex}'})

        elif action == 'editcontactoemergencia':
            try:
                f = ContactoEmergenciaForm(request.POST)
                administrativo = Persona.objects.get(id=request.GET['id'])
                if f.is_valid():
                    datosextension = administrativo.datos_extension()
                    datosextension.contactoemergencia=f.cleaned_data['contactoemergencia']
                    datosextension.parentescoemergencia=f.cleaned_data['parentescoemergencia']
                    datosextension.telefonoemergencia=f.cleaned_data['telefonoemergencia']
                    datosextension.correoemergencia=f.cleaned_data['correoemergencia']
                    datosextension.telefonoconvemergencia=f.cleaned_data['telefonoconvemergencia']
                    datosextension.save(request)
                    log(u'editó contacto de emergencia: %s' % datosextension, request, "edit")
                    return JsonResponse({'result': False, 'mensaje': 'Guardado con exito'})
                else:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()],
                                         "mensaje": "Error en el formulario"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', 'mensaje': f'{ex}'})


        elif action == 'consul_contrato':
            try:
                personacontrato = PersonaContratos.objects.get(pk=request.POST['idcontra'], status=True)
                numero = personacontrato.numerodocumento
                relacionada = personacontrato.contratacionrelacionada
                return JsonResponse({"result": "ok","numero": numero,"relacionada": relacionada })
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": "Error al eliminar los datos."})

        if action == 'consul_accion':
            try:
                personaaccion = PersonaAcciones.objects.get(pk=request.POST['idaccion'], status=True)
                numero = personaaccion.numerodocumento
                tipo = personaaccion.tipo.nombre
                return JsonResponse({"result": "ok","numero": numero,"tipo": tipo })
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": "Error al eliminar los datos."})


        if action == 'tutoriaaprobado':
            try:
                tutoria = CertificadoTutoriaHV.objects.get(pk=int(request.POST['id']))
                tutoria.aprobado = True if request.POST['valor'] == 'y' else False
                tutoria.save(request)
                log(u'Verifico certificado de tutoria: %s' % tutoria, request, "edit")
                return JsonResponse({'result': 'ok'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', 'mensaje': u'Error al guardar los datos'})

        if action == 'verificaridioma':
            try:
                idioma = IdiomaDomina.objects.get(pk=int(request.POST['id']))
                idioma.validado = True if request.POST['valor'] == 'y' else False
                idioma.save(request)
                log(u'Verifico idioma: %s' % idioma, request, "edit")
                return JsonResponse({'result': 'ok'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', 'mensaje': u'Error al guardar los datos'})

        if action == 'addarchivoidioma':
            try:
                f = ArchivoIdiomaForm(request.POST, request.FILES)
                if f.is_valid():
                    personaadmin = Persona.objects.get(pk=int(request.POST['ida']))
                    titulacion = IdiomaDomina.objects.get(pk=int(request.POST['id']))
                    newfile = request.FILES['archivo']
                    newfile._name = generar_nombre("idioma_", newfile._name)
                    titulacion.archivo = newfile
                    titulacion.save(request)
                    log(u'Adiciono archivo de idioma: %s' % personaadmin, request, "add")
                    return JsonResponse({'result': 'ok'})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', 'mensaje': u'Error al guardar los datos'})

        if action == 'addarchivoexperiencia':
            try:
                f = ArchivoExperienciaForm(request.POST, request.FILES)
                if f.is_valid():
                    personaadmin = Persona.objects.get(pk=int(request.POST['ida']))
                    titulacion = ExperienciaLaboral.objects.get(pk=int(request.POST['id']))
                    newfile = request.FILES['archivo']
                    newfile._name = generar_nombre("experiencialaboral_", newfile._name)
                    titulacion.archivo = newfile
                    titulacion.save(request)
                    log(u'Adiciono archivo de referncia laboral: %s' % persona, request, "add")
                    return JsonResponse({'result': 'ok'})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', 'mensaje': u'Error al guardar los datos'})




        if action == 'addarchivocapacitacion':
            try:
                personaadmin = Persona.objects.get(pk=int(request.POST['ida']))
                f = ArchivoTitulacionForm(request.POST, request.FILES)
                if f.is_valid():
                    titulacion = Capacitacion.objects.get(pk=int(request.POST['id']))
                    newfile = request.FILES['archivo']
                    newfile._name = generar_nombre("capacitacion_", newfile._name)
                    titulacion.archivo = newfile
                    titulacion.save(request)
                    log(u'Adiciono archivo de capacitacion: %s' % personaadmin, request, "add")
                    return JsonResponse({'result': 'ok'})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', 'mensaje': u'Error al guardar los datos'})

        if action == 'rechazarcontratofirmado':
            try:
                historial = HistorialArchivosContratos.objects.get(pk=int(request.POST['id']))
                historial.estado_archivo = 4
                historial.save()
                log(u'Rechazó contrato firmado del personal laboral: %s' % persona, request, "del")

                return JsonResponse({'result': 'ok'})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})


        if action == 'addidioma':
            try:
                form = IdiomaDominaForm(request.POST)
                if form.is_valid():
                    personaadmin = Persona.objects.get(pk=int(request.POST['id']))
                    idioma = IdiomaDomina(persona=personaadmin,
                                          idioma=form.cleaned_data['idioma'],
                                          lectura=form.cleaned_data['lectura'],
                                          escritura=form.cleaned_data['escritura'],
                                          oral=form.cleaned_data['oral'])
                    idioma.save(request)
                    if 'archivo' in request.FILES:
                        newfile = request.FILES['archivo']
                        newfile._name = generar_nombre("idioma_", newfile._name)
                        idioma.archivo = newfile
                        idioma.save(request)
                    log(u'Adiciono idioma que domina: %s' % idioma, request, "add")
                    return JsonResponse({'result': 'ok'})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', 'mensaje': u'Error al guardar los datos'})


        if action == 'editidioma':
            try:
                form = IdiomaDominaForm(request.POST, request.FILES)
                if form.is_valid():
                    idioma = IdiomaDomina.objects.get(pk=request.POST['id'])
                    idioma.idioma = form.cleaned_data['idioma']
                    idioma.lectura = form.cleaned_data['lectura']
                    idioma.escritura = form.cleaned_data['escritura']
                    idioma.oral = form.cleaned_data['oral']
                    idioma.save(request)
                    if 'archivo' in request.FILES:
                        newfile = request.FILES['archivo']
                        newfile._name = generar_nombre("idioma_", newfile._name)
                        idioma.archivo = newfile
                        idioma.save(request)
                    log(u'Modifico idioma que domina: %s' % idioma, request, "add")
                    return JsonResponse({'result': 'ok'})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', 'mensaje': u'Error al guardar los datos datos'})

        if action == 'delidioma':
            try:
                idioma = IdiomaDomina.objects.get(pk=request.POST['id'])
                log(u"Elimino idioma que domina: %s" % idioma, request, "del")
                idioma.delete()
                return JsonResponse({'result': 'ok'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', 'mensaje': u'Error al eliminar los datos'})

        if action == 'detallearticulo':
            try:
                data['articulo'] = articulos = ArticuloInvestigacion.objects.get(pk=request.POST['id'])
                data['evidencias'] = Evidencia.objects.filter(status=True, matrizevidencia_id=3)
                data['formevidencias'] = EvidenciaForm()
                # data['evidenciasarticulo'] = evidencias = DetalleEvidencias.objects.filter(articulo=articulos)
                template = get_template("th_hojavida/detallearticulo.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'html': json_content})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        if action == 'detalleponencia':
            try:
                data['ponencias'] = ponencias = PonenciasInvestigacion.objects.get(pk=request.POST['id'])

                data['evidencias'] = Evidencia.objects.filter(status=True, matrizevidencia_id=4)
                data['formevidencias'] = EvidenciaForm()
                template = get_template("th_hojavida/detalleponencia.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'html': json_content})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        if action == 'detalleexplicacion':
            try:
                if 'id' in request.POST:
                    data['accion'] = PersonaAcciones.objects.get(pk=int(request.POST['id']))
                    template = get_template("th_personal/detalleaccionpersonalexplicacion.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'html': json_content})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        if action == 'detalleexplicacion2':
            try:
                if 'id' in request.POST:
                    data['accion'] = AccionPersonal.objects.get(pk=int(request.POST['id']))
                    template = get_template("th_personal/detalleaccionpersonalexplicacion.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'html': json_content})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        if action == 'detallecontrato':
            try:
                if 'id' in request.POST:
                    data['contrato'] = PersonaContratos.objects.get(pk=int(request.POST['id']))
                    template = get_template("th_personal/detallecontrato.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'html': json_content})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'addarchivocontratofirmadodirector':
            try:
                with transaction.atomic():
                    persona = request.session['persona']
                    contrato = PersonaContratos.objects.get(pk=int(request.POST['id']))
                    if int(request.POST['estado_archivo']) == 2 or int(request.POST['estado_archivo']) == 5 or int(request.POST['estado_archivo']) == 3:
                        if not 'archivo' in request.FILES:
                            return JsonResponse({"result": True, "mensaje": "Subir contrato firmado en formato .pdf"}, safe=False)

                    if 'archivo' in request.FILES:
                        d = request.FILES['archivo']
                        newfilesd = d._name
                        ext = newfilesd[newfilesd.rfind("."):]
                        if ext == '.pdf':
                            a = 1
                        else:
                            return JsonResponse({'result': True, "message": 'Error, solo archivos .pdf.'}, safe=False)
                        if d.size > 12582912:
                            return JsonResponse({'result': True, "message": 'Error, archivo mayor a 12 Mb.'}, safe=False)

                    f = ArchivoHistorialContratoDirectorForm(request.POST, request.FILES)
                    if f.is_valid():
                        historial = HistorialArchivosContratos(personacontrato=contrato,
                                                               estado_archivo=f.cleaned_data['estado_archivo'],
                                                               observacion=f.cleaned_data['observacion']
                                                               )
                        historial.save(request)
                        newfile = None
                        if 'archivo' in request.FILES:
                            newfile = request.FILES['archivo']
                            newfile._name = generar_nombre("contratopersonalfirmadodirector", newfile._name)
                            historial.archivo = newfile
                            historial.save(request)
                        contrato.ultimo_archivo = historial.id
                        contrato.save(request)
                        log(u'Adiciono contratro firmado por director: %s' % persona, request, "add")
                        return JsonResponse({"result": False}, safe=False)
                    else:
                        return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()], "message": "Error en el formulario"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        elif action == 'editarchivocontratofirmadodirector':
            try:
                with transaction.atomic():
                    persona = request.session['persona']
                    if not 'archivo' in request.FILES:
                        return JsonResponse({"result": True, "mensaje": "Subir contrato firmado en formato .pdf"}, safe=False)

                    if 'archivo' in request.FILES:
                        d = request.FILES['archivo']
                        newfilesd = d._name
                        ext = newfilesd[newfilesd.rfind("."):]
                        if ext == '.pdf':
                            a = 1
                        else:
                            return JsonResponse({'result': True, "message": 'Error, solo archivos .pdf.'}, safe=False)
                        if d.size > 12582912:
                            return JsonResponse({'result': True, "message": 'Error, archivo mayor a 12 Mb.'}, safe=False)

                    f = ArchivoHistorialContratoForm(request.POST, request.FILES)
                    if f.is_valid():
                        newfile = None
                        if 'archivo' in request.FILES:
                            newfile = request.FILES['archivo']
                            newfile._name = generar_nombre("contratopersonalfirmadodirector", newfile._name)
                            historial = HistorialArchivosContratos.objects.get(pk=int(request.POST['id']))
                            historial.archivo = newfile
                            historial.save(request)
                        log(u'Editó contratro frimado por director: %s' % persona, request, "edit")
                        return JsonResponse({"result": False}, safe=False)
                    else:
                        return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()], "message": "Error en el formulario"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)


        elif action == 'adddocumentofirmado':
            try:
                with transaction.atomic():
                    newfile = None
                    form = AccionPersonalDocumentoForm(request.POST, request.FILES)
                    if form.is_valid():
                        accionpersonal = PersonaAcciones.objects.get(pk=int(request.POST['id']))
                        if 'archivofirmado' in request.FILES:
                            newfile = request.FILES['archivofirmado']
                            extension = newfile._name.split('.')
                            tam = len(extension)
                            exte = extension[tam - 1]
                            if newfile.size > 4194304:
                                return JsonResponse(
                                    {"result": True, "mensaje": u"Error, el tamaño del archivo es mayor a 4 Mb."})
                            if exte in ['pdf', 'jpg', 'jpeg', 'png', 'jpeg', 'peg']:
                                newfile._name = generar_nombre("AccionPersonalFirmado_", newfile._name)
                            else:
                                return JsonResponse(
                                    {"result": True, "mensaje": u"Error, solo archivos .pdf,.jpg, .jpeg"})
                            # accionpersonal.archivofirmado = newfile
                        historialdocumento = HistoricoDocumentosPersonaAcciones(personaaccion=accionpersonal, archivofirmado=newfile)
                        historialdocumento.save(request)
                        accionpersonal.estadoarchivo = 2
                        accionpersonal.save(request)
                        log(u'Adiciono Documento firmado en Accion de personal: %s' % accionpersonal, request, "add")
                        return redirect('/th_personal?action=detallepersonal&ida={}'.format(accionpersonal.persona.pk))
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({"result": True, "mensaje": "Complete los datos requeridos."}, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        elif action == 'descargarcomprimido':
            with transaction.atomic():
                try:
                    form = DescargarCompromidoForm(request.POST)
                    if form.is_valid():
                        data['periodo']=periodo=form.cleaned_data['periodo']
                        titulo = f'Generación de archivo .zip de documentos de familiares pertenecientes al periodo {periodo.anio} en proceso.'
                        noti = Notificacion(cuerpo='Se inicializo la compresion de documentos de familiares que registran gastos.',
                                            titulo=titulo, destinatario=persona,
                                            url='',
                                            prioridad=1, app_label='SGA-SAGEST',
                                            fecha_hora_visible=datetime.now() + timedelta(days=1), tipo=2,
                                            en_proceso=True)
                        noti.save(request)
                        descarga_masica_documentos_familiares_background(request=request, data=data, notif=noti.pk).start()
                        messages.success(request,'Generando archivo .zip')
                        return JsonResponse({"result": False, "mensaje": f'Generando archivo .zip'})
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({"result": True, "mensaje": "Complete los datos requeridos."})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": True, "mensaje": f"{ex}"})


        elif action == 'repcargas':
            with transaction.atomic():
                try:
                    form = DescagarCargasFechasForm(request.POST)
                    if form.is_valid():
                        data['periodo'] = periodo = PeriodoGastosPersonales.objects.get(pk=encrypt(request.POST['id']))
                        data['fechainicio'] = form.cleaned_data['fechainicio']
                        data['fechafin'] = form.cleaned_data['fechafin']
                        data['todos'] = form.cleaned_data['todos']
                        titulo = f'Generación de reporte de cargas familiares pertenecientes al periodo {periodo.anio} en proceso.'
                        noti = Notificacion(cuerpo='Se inició la descarga del reporte, por favor espere.',
                                            titulo=titulo, destinatario=persona,
                                            url='',
                                            prioridad=1, app_label='SGA-SAGEST',
                                            fecha_hora_visible=datetime.now() + timedelta(days=1), tipo=2,
                                            en_proceso=True)
                        noti.save(request)
                        descarga_cargas_gastosp_background(request=request, data=data, notif=noti.pk).start()
                        messages.success(request,'Generando reporte')
                        return JsonResponse({"result": False, "mensaje": f'Generando reporte'})
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({"result": True, "mensaje": "Complete los datos requeridos."})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": True, "mensaje": f"{ex}"})

        elif action == 'importardecimo':
            try:
                f = SagestImportarXLSForm(request.POST)
                if f.is_valid():
                    nfile = request.FILES['archivo']
                    archivo = io.BytesIO(nfile.read())
                    workbook = load_workbook(filename=archivo, read_only=False)
                    aimporta = workbook[workbook.sheetnames[0]]
                    cidentificacion = 1
                    cmodalidad = 3
                    for rowx in range(2, aimporta.max_row + 1):
                        identificacion = aimporta.cell(row=rowx, column=cidentificacion).value
                        modalidad = aimporta.cell(row=rowx, column=cmodalidad).value.upper().strip()
                        distributivo = Persona.objects.filter(
                            (Q(cedula__icontains=str(identificacion)) | Q(pasaporte__icontains=identificacion)), status=True)
                        seleccion= 1 if modalidad =='ACUMULA' else 2
                        if distributivo:
                            distributivo=distributivo[0]
                            registrobase = RegistroDecimo.objects.filter(persona=distributivo,status=True).order_by('id').last()
                            if not RegistroDecimo.objects.filter(persona=distributivo,status=True, activo=True,seleccion=seleccion).exists():
                                if registrobase:
                                    registrobase.activo=False
                                    registrobase.fechafin=datetime.now().date()
                                    registrobase.save(update_fields=['activo','fechafin'])
                                registro = RegistroDecimo(persona=distributivo,
                                                          estado=2,
                                                          seleccion=seleccion,
                                                          fechainicio=datetime.now().date(),
                                                          activo=True
                                                          )
                                registro.save(request)
                                log('%s importó masivamente el registro de décimo de %s con tipo %s' % (
                                    persona.nombre_completo_minus(), distributivo.nombre_completo_minus(),modalidad),request,"add")

                    return JsonResponse({"result": False}, safe=False)
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, 'mensaje': str(ex)}, safe=False)

        elif action == 'importarpersona':
            try:
                periodoth = ConfiguraPeriodotthh.objects.get(id=encrypt_id(request.POST['id']))
                f = SagestImportarXLSForm(request.POST)
                if f.is_valid():
                    nfile = request.FILES['archivo']
                    archivo = io.BytesIO(nfile.read())
                    workbook = load_workbook(filename=archivo, read_only=False)
                    aimporta = workbook[workbook.sheetnames[0]]
                    cidentificacion = 1
                    cmodalidad = 3
                    for rowx in range(2, aimporta.max_row + 1):
                        identificacion = aimporta.cell(row=rowx, column=cidentificacion).value
                        distributivo = Persona.objects.filter((Q(cedula__icontains=str(identificacion)) | Q(pasaporte__icontains=identificacion)), status=True).first()
                        if distributivo:
                            # registrobase = distributivo.personaperiodotthh_set.filter(status=True).order_by('id').last()
                            if not PersonaPeriodotthh.objects.filter(persona=distributivo,periodotthh=periodoth, status=True).exists():
                                # if registrobase:
                                #     registrobase.activo=False
                                #     registrobase.fechafin=datetime.now().date()
                                #     registrobase.save(update_fields=['activo','fechafin'])
                                registro = PersonaPeriodotthh(periodotthh=periodoth,
                                                              persona=distributivo)
                                registro.save(request)
                                log('%s importó masivamente el registro de personas al periodo tthh de %s' % (persona.nombre_completo_minus(), distributivo.nombre_completo_minus()),request,"add")

                    return JsonResponse({"result": False}, safe=False)
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, 'mensaje': str(ex)}, safe=False)

        elif action == 'addperiodotthh':
            try:
                form = PeriodoTTHHForm(request.POST)
                if not form.is_valid():
                    transaction.set_rollback(True)
                    form_error = [{k: v[0]} for k, v in form.errors.items()]
                    return JsonResponse({'result': True, "form": form_error, "mensaje": "Error en el formulario"})
                periodoth = ConfiguraPeriodotthh(nombre=form.cleaned_data['nombre'],
                                                 fechainicio=form.cleaned_data['fechainicio'],
                                                 fechafin=form.cleaned_data['fechafin'])
                periodoth.save(request)
                log(f'Agrego requisito periodo tthh: {periodoth}', request, 'add')
                return JsonResponse({'result': False, 'mensaje': 'Guardado con éxito'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': str(ex)})

        elif action == 'editperiodotthh':
            try:
                instancia = ConfiguraPeriodotthh.objects.get(id=encrypt_id(request.POST['id']))
                form = PeriodoTTHHForm(request.POST, request.FILES, instancia=instancia)
                if not form.is_valid():
                    transaction.set_rollback(True)
                    form_error = [{k: v[0]} for k, v in form.errors.items()]
                    return JsonResponse({'result': True, "form": form_error, "mensaje": "Error en el formulario"})
                instancia.nombre = form.cleaned_data['nombre']
                instancia.fechainicio = form.cleaned_data['fechainicio']
                instancia.fechafin = form.cleaned_data['fechafin']
                instancia.activo = form.cleaned_data['activo']
                instancia.save(request)
                log(f'Edito periodo tthh: {instancia}', request, 'edit')
                return JsonResponse({'result': False, 'mensaje': 'Guardado con éxito'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': str(ex)})

        elif action == 'addrequisitoperiodotthh':
            try:
                form = RequisitoPeriodoTTHHForm(request.POST, request.FILES)
                if not form.is_valid():
                    transaction.set_rollback(True)
                    form_error = [{k: v[0]} for k, v in form.errors.items()]
                    return JsonResponse({'result': True, "form": form_error, "mensaje": "Error en el formulario"})
                requisito = RequisitoPeriodotthh(periodotthh_id=encrypt_id(request.POST['id']),
                                                 nombre=form.cleaned_data['nombre'],
                                                 descripcion=form.cleaned_data['descripcion'],
                                                 link=form.cleaned_data['link'],
                                                 mostrar=form.cleaned_data['mostrar'],
                                                 opcional=form.cleaned_data['opcional'],
                                                 archivo=form.cleaned_data['archivo'])
                requisito.save(request)
                log(f'Agrego requisito periodo tthh: {requisito}', request, 'add')
                return JsonResponse({'result': False, 'mensaje': 'Guardado con éxito'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': str(ex)})

        elif action == 'editrequisitoperiodotthh':
            try:
                instancia = RequisitoPeriodotthh.objects.get(id=encrypt_id(request.POST['id']))
                form = RequisitoPeriodoTTHHForm(request.POST, request.FILES, instancia=instancia)
                if not form.is_valid():
                    transaction.set_rollback(True)
                    form_error = [{k: v[0]} for k, v in form.errors.items()]
                    return JsonResponse({'result': True, "form": form_error, "mensaje": "Error en el formulario"})
                instancia.nombre = form.cleaned_data['nombre']
                instancia.descripcion = form.cleaned_data['descripcion']
                instancia.link = form.cleaned_data['link']
                instancia.mostrar = form.cleaned_data['mostrar']
                instancia.opcional = form.cleaned_data['opcional']
                instancia.archivo = form.cleaned_data['archivo']
                instancia.save(request)
                log(f'Edito requisito periodo tthh: {instancia}', request, 'edit')
                return JsonResponse({'result': False, 'mensaje': 'Guardado con éxito'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': str(ex)})

        elif action == 'delrequisito':
            try:
                requisito = RequisitoPeriodotthh.objects.get(id=encrypt_id(request.POST['id']))
                requisito.status = False
                requisito.save(request)
                log(f'Elimino requisito periodo tthh: {requisito}', request, 'del')
                res_json = {"error": False}
            except Exception as ex:
                res_json = {'error': True, "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)

        elif action == 'validarrequisitotthh':
            with transaction.atomic():
                try:
                    data['hoy'] = hoy = datetime.now()
                    instance = DocumentoPersonaPeriodotthh.objects.get(pk=int(request.POST['id']))
                    form = ValidarRequisitoTTHHForm(request.POST)
                    if not form.is_valid():
                        transaction.set_rollback(True)
                        return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()],
                                             "mensaje": "Error en el formulario"})
                    instance.estados = int(form.cleaned_data['estado'])
                    instance.observacion = form.cleaned_data['observacion']
                    instance.f_validacion = hoy
                    instance.save(request)
                    personaperiodotthh = instance.personaperiodotthh
                    documentos = DocumentoPersonaPeriodotthh.objects.filter(personaperiodotthh=personaperiodotthh, status=True).exclude(id=instance.id)
                    mensaje = ''
                    if instance.estados == 1 and personaperiodotthh.doc_validacion() == 1 and personaperiodotthh.estado_requisito != 1:
                        personaperiodotthh.estado_requisito = 1
                        personaperiodotthh.save(request)
                        mensaje = f'Los documentos subidos fueron aprobados requisitos de ingreso'
                    elif personaperiodotthh.doc_validacion() == 0 and personaperiodotthh.estado_requisito != 0:
                        personaperiodotthh.estado_requisito = 0
                        personaperiodotthh.save(request)
                    elif instance.estados == 2:
                        personaperiodotthh.estado_requisito = 2
                        personaperiodotthh.save(request)
                        mensaje = f'Documento {instance} pendiente de corregir.'
                    elif instance.estados == 4:
                        mensaje = f'Documento {instance} fue rechazado.'

                    if instance.estados == 1 and personaperiodotthh.doc_validacion() == 1 or instance.estados != 1:
                        titulo = u"Validación de documentos cargados en requisitos de ingreso de {} - ({})".format(personaperiodotthh, instance.get_estados_display())
                        notificacion(titulo,
                                     mensaje, personaperiodotthh.persona, None, f'/th_hojavida?action=requisitosingreso',
                                     instance.pk, 1, 'sga-sagest', DocumentoPersonaPeriodotthh, request)

                    diccionario = {'id': instance.id, 'observacion': instance.observacion, 'estado': instance.get_estados_display(), 'idestado': instance.estados, 'color': instance.color_estado()}
                    log(u'Valido documento de requisitos de ingreso: %s' % instance, request, "edit")
                    return JsonResponse({'result': True, 'data_return': True, 'mensaje': u'Guardado con exito', 'data': diccionario}, safe=False)
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": True, "mensaje": "Intentelo más tarde.".format(str(ex))}, safe=False)

        elif action == 'delperiodotthh':
            try:
                periodotthh = ConfiguraPeriodotthh.objects.get(id=encrypt_id(request.POST['id']))
                log(f'Elimino periodo tthh: {periodotthh}', request, 'del')
                periodotthh.delete()
                res_json = {"error": False}
            except Exception as ex:
                res_json = {'error': True, "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)

        elif action == 'delpersonaperiodo':
            try:
                requisito = PersonaPeriodotthh.objects.get(id=encrypt_id(request.POST['id']))
                log(f'Elimino persona de periodo tthh: {requisito.persona}', request, 'del')
                requisito.delete()
                res_json = {"error": False}
            except Exception as ex:
                res_json = {'error': True, "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)

        elif action == 'cambiarestado':
            try:
                tipo = request.POST.get('args', '')
                requisito = RequisitoPeriodotthh.objects.get(id=encrypt_id(request.POST['id']))
                if tipo == 'mostrar':
                    requisito.mostrar = eval(request.POST['val'].capitalize())
                    requisito.save(request)
                elif tipo == 'opcional':
                    requisito.opcional = eval(request.POST['val'].capitalize())
                    requisito.save(request)
                log(f'Edito requisito: {requisito}', request, 'edit')
                return JsonResponse({'result': False, 'mensaje': 'Guardado con éxito'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': str(ex)})

        elif action == 'descargardecimo':
            with transaction.atomic():
                try:
                    titulo = f'Generación de archivo .zip de actas firmadas de décimo en proceso.'
                    noti = Notificacion(cuerpo='Se inicializo la compresion de actas firmadas de décimo.',
                                        titulo=titulo, destinatario=persona,
                                        url='',
                                        prioridad=1, app_label='SGA-SAGEST',
                                        fecha_hora_visible=datetime.now() + timedelta(days=1), tipo=2,
                                        en_proceso=True)
                    noti.save(request)
                    descarga_masiva_actas_firmadas_background(request=request, data=data, notif=noti.pk).start()
                    messages.success(request, 'Generando archivo .zip')
                    return JsonResponse({"result": 'ok', "mensaje": f'Generando archivo .zip'})

                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": 'bad', "mensaje": f"{ex}"})

        elif action == 'descargarcomprimidomasivo':
            with transaction.atomic():
                try:
                    estado, requisito, filtro= request.POST.get('estado', ''), \
                                               request.POST.getlist('requisito', []), \
                                               Q(status=True)
                    # if requisito:
                    #     filtro = filtro & Q(documentopersonaperiodotthh__requisito_id__in=requisito)
                    if estado:
                        filtro = filtro & Q(estado_requisito=int(estado))
                    data['ePersonasPeriodo'] = ePersonasPeriodo = PersonaPeriodotthh.objects.filter(filtro).exclude(documentopersonaperiodotthh__isnull=True)
                    if not ePersonasPeriodo:
                        raise NameError('No existen registros con la configuración de descarga ingresada')
                    titulo = f'Generación de archivo .zip de documentos de requisitos de ingreso'
                    noti = Notificacion(cuerpo='Se inicializo la compresión de documentos de requisitos de ingreso cargados al sistema',
                                        titulo=titulo, destinatario=persona,
                                        url='',
                                        prioridad=1, app_label='SGA-SAGEST',
                                        fecha_hora_visible=datetime.now() + timedelta(days=1), tipo=2,
                                        en_proceso=True)
                    noti.save(request)
                    descarga_masiva_requisitos_ingreso_background(request=request, data=data, notif=noti.pk).start()
                    return JsonResponse({"result": False, "mensaje": u"Generando archivo .zip"})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": True, "mensaje": f"{ex}"})

        # elif action == 'repcargas':
        #     with transaction.atomic():
        #         try:
        #             form = DescagarCargasFechasForm(request.POST)
        #             if form.is_valid():
        #                 periodo = PeriodoGastosPersonales.objects.get(pk=encrypt(request.POST['id']))
        #                 GastosPersonales.objects.values_list('persona_id').filter(status=True,
        #                                                                           periodogastospersonales=periodo)
        #                 ids = GastosPersonales.objects.values_list('persona_id').filter(status=True,
        #                                                                                 periodogastospersonales=periodo)
        #                 filtro = Q(status=True, actagenerada=True,aplicaproyeccion=True)
        #                 if not form.cleaned_data['todos']:
        #                     filtro = Q(status=True, actagenerada=True, aplicaproyeccion=True,
        #                                fecha_modificacion__lte=form.cleaned_data['fechainicio'],
        #                                fecha_modificacion__gte=form.cleaned_data['fechafin'])
        #
        #
        #                 __author__ = 'Unemi'
        #                 output = io.BytesIO()
        #                 workbook = xlsxwriter.Workbook(output)
        #                 ws = workbook.add_worksheet('cargas')
        #                 ws.set_column(0, 20, 50)
        #                 formatoceldagris = workbook.add_format(
        #                     {'align': 'center', 'border': 1, 'text_wrap': True, 'fg_color': '#B6BFC0'})
        #                 formatoceldaleft = workbook.add_format({'text_wrap': True, 'align': 'left'})
        #                 ws.write('A1', 'IDENTIFICACIÓN', formatoceldagris)
        #                 ws.write('B1', 'SERVIDOR', formatoceldagris)
        #                 ws.write('C1', 'CANT. CARGAS', formatoceldagris)
        #                 i = 2
        #
        #                 personas = Persona.objects.filter(status=True, id__in=ids)
        #                 for per in personas:
        #                     familiares = per.personadatosfamiliares_set.filter(filtro)
        #                     ws.write('A%s' % i, str(per.identificacion()), formatoceldaleft)
        #                     ws.write('B%s' % i, str(per.nombre_completo_inverso_2()), formatoceldaleft)
        #                     ws.write('C%s' % i, str(familiares.count()), formatoceldaleft)
        #                     i += 1
        #                 workbook.close()
        #                 output.seek(0)
        #                 filename = 'reporte_cargas_familiares.xlsx'
        #                 response = HttpResponse(output,
        #                                         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        #                 response['Content-Disposition'] = 'attachment; filename=%s' % filename
        #                 return response
        #             else:
        #                 transaction.set_rollback(True)
        #                 return JsonResponse({"result": True, "mensaje": "Complete los datos requeridos."})
        #         except Exception as ex:
        #             transaction.set_rollback(True)
        #             return JsonResponse({"result": True, "mensaje": f"{ex}"})
        return JsonResponse({"result": "bad", "mensaje": u"Solicitud Incorrecta."})
    else:
        if 'action' in request.GET:
            data['action']= action= request.GET['action']
            if 'ida' in request.GET:
                data['administrativo'] = administrativo = Persona.objects.get(pk=int(request.GET['ida']))

            if action == 'datospersonales':
                try:
                    data['title'] = u'Datos personales'
                    data['personaadmin'] = personaadmin = Persona.objects.get(pk=int(request.GET['id']))
                    maternidad = personaadmin.personadetallematernidad_set.filter(Q(status_gestacion=True) | Q(status_lactancia=True)).first()
                    if personaadmin.sexo:
                        if personaadmin.sexo.id == 1 and maternidad:
                            data['form_2'] = PersonaDetalleMaternidad(initial={'estadogestacion': maternidad.status_gestacion,
                                                                               'semanasembarazo': maternidad.semanasembarazo,
                                                                               'lactancia': maternidad.status_lactancia,
                                                                               'fechaparto': maternidad.fechaparto})
                        elif not maternidad:
                            data['form_2'] = PersonaDetalleMaternidad()
                    form = DatosPersonalesForm(initial={'nombres': personaadmin.nombres,
                                                        'apellido1': personaadmin.apellido1,
                                                        'apellido2': personaadmin.apellido2,
                                                        'cedula': personaadmin.cedula,
                                                        'pasaporte': personaadmin.pasaporte,
                                                        'sexo': personaadmin.sexo,
                                                        'anioresidencia': personaadmin.anioresidencia,
                                                        'nacimiento': personaadmin.nacimiento,
                                                        'nacionalidad': personaadmin.nacionalidad,
                                                        'email': personaadmin.email,
                                                        'estadocivil': personaadmin.estado_civil(),
                                                        'libretamilitar': personaadmin.libretamilitar})
                    form.editar(sinnombres=False)
                    data['form'] = form
                    return render(request, "th_personal/datospersonales.html", data)
                except:
                    pass

            elif action == 'datosnacimiento':
                try:
                    data['title'] = u'Datos de nacimiento'
                    data['personaadmin'] = personaadmin = Persona.objects.get(pk=int(request.GET['id']))
                    form = DatosNacimientoForm(initial={'paisnacimiento': personaadmin.paisnacimiento,
                                                        'provincianacimiento': personaadmin.provincianacimiento,
                                                        'cantonnacimiento': personaadmin.cantonnacimiento,
                                                        'parroquianacimiento': personaadmin.parroquianacimiento})
                    form.editar(personaadmin)
                    data['form'] = form
                    return render(request, "th_personal/datosnacimiento.html", data)
                except Exception as ex:
                    pass

            elif action == 'rephijos':
                try:
                    __author__ = 'Unemi'
                    output = io.BytesIO()
                    workbook = xlsxwriter.Workbook(output)
                    ws = workbook.add_worksheet('diagnostico')
                    ws.set_column(0, 20, 50)
                    formatoceldagris = workbook.add_format(
                        {'align': 'center', 'border': 1, 'text_wrap': True, 'fg_color': '#B6BFC0'})
                    formatoceldaleft = workbook.add_format({'text_wrap': True, 'align': 'left'})
                    ws.write('A1', 'IDENTIFICACIÓN', formatoceldagris)
                    ws.write('B1', 'SERVIDOR', formatoceldagris)
                    ws.write('C1', 'CARGO', formatoceldagris)
                    ws.write('D1', 'ESTADO CIVIL', formatoceldagris)
                    ws.write('E1', 'SEXO', formatoceldagris)
                    ws.write('F1', 'EMBARAZO', formatoceldagris)
                    ws.write('G1', 'PARENTESCO', formatoceldagris)
                    ws.write('H1', 'NOMBRE DE FAMILIAR', formatoceldagris)
                    ws.write('I1', 'IDENTIFICACIÓN', formatoceldagris)
                    ws.write('J1', 'FECHA DE NACIMIENTO', formatoceldagris)
                    ws.write('K1', 'EDAD EN MESES', formatoceldagris)
                    ws.write('L1', 'APLICA PROYECCIÓN', formatoceldagris)
                    ws.write('M1', 'ACTA GENERADA', formatoceldagris)
                    ws.write('N1', 'ES SERVIDOR PÚBLICO', formatoceldagris)
                    ws.write('O1', 'TIENE CUSTODIA', formatoceldagris)
                    ws.write('P1', 'CENTRO DE CUIDADO', formatoceldagris)
                    ws.write('Q1', 'SEXO', formatoceldagris)
                    i = 2
                    familiares = PersonaDatosFamiliares.objects.filter(personafamiliar__isnull=False,persona_id__in=(DistributivoPersona.objects.values_list('persona_id').filter(status=True))).distinct()
                    for familiar in familiares:
                        ws.write('A%s' % i, str(familiar.persona.identificacion()), formatoceldaleft)
                        ws.write('B%s' % i, str(familiar.persona.nombre_completo_inverso_2()), formatoceldaleft)
                        ws.write('C%s' % i, str(familiar.persona.mi_cargo_actual()), formatoceldaleft)
                        ws.write('D%s' % i, str(familiar.persona.estado_civil()), formatoceldaleft)
                        ws.write('E%s' % i, str(familiar.persona.sexo), formatoceldaleft)
                        ws.write('F%s' % i, str(familiar.persona.maternidad().semanas_gestacion()), formatoceldaleft)
                        ws.write('G%s' % i, str(familiar.parentesco), formatoceldaleft)
                        ws.write('H%s' % i, str(familiar.personafamiliar.nombre_completo_inverso_2()), formatoceldaleft)
                        ws.write('I%s' % i, str(familiar.personafamiliar.identificacion()), formatoceldaleft)
                        ws.write('J%s' % i, str(familiar.nacimiento), formatoceldaleft)
                        ws.write('K%s' % i, str(familiar.edad_meses()), formatoceldaleft)
                        ws.write('L%s' % i, 'SI' if familiar.aplicaproyeccion else '', formatoceldaleft)
                        ws.write('M%s' % i, 'SI' if familiar.actagenerada else '', formatoceldaleft)
                        ws.write('N%s' % i, 'SI' if familiar.esservidorpublico else '', formatoceldaleft)
                        ws.write('O%s' % i, 'SI' if familiar.bajocustodia else '', formatoceldaleft)
                        ws.write('P%s' % i, str(familiar.get_centrocuidado_display()), formatoceldaleft)
                        ws.write('Q%s' % i, str(familiar.personafamiliar.sexo), formatoceldaleft)

                        i += 1
                    workbook.close()
                    output.seek(0)
                    filename = 'reporte_cargas_familiares.xlsx'
                    response = HttpResponse(output,
                                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    response['Content-Disposition'] = 'attachment; filename=%s' % filename
                    return response
                except Exception as ex:
                    pass

            elif action == 'repencuesta':
                try:
                    __author__ = 'Unemi'
                    output = io.BytesIO()
                    workbook = xlsxwriter.Workbook(output)
                    ws = workbook.add_worksheet('resultados')
                    ws.set_column(0, 20, 50)
                    formatoceldagris = workbook.add_format(
                        {'align': 'center', 'border': 1, 'text_wrap': True, 'fg_color': '#B6BFC0'})
                    formatoceldaleft = workbook.add_format({'text_wrap': True, 'align': 'left'})
                    ws.write('A1', 'IDENTIFICACIÓN', formatoceldagris)
                    ws.write('B1', 'SERVIDOR', formatoceldagris)
                    ws.write('C1', 'CARGO', formatoceldagris)
                    ws.write('D1', 'ESTADO CIVIL', formatoceldagris)
                    ws.write('E1', 'SEXO', formatoceldagris)
                    ws.write('F1', 'REGIMEN LABORAL', formatoceldagris)
                    ws.write('G1', 'UNIDAD', formatoceldagris)
                    ws.write('H1', 'TIENE HIJOS', formatoceldagris)
                    ws.write('I1', 'NRO HIJOS', formatoceldagris)
                    i = 2
                    personas = PersonaExtension.objects.filter(persona_id__in=(DistributivoPersona.objects.values_list('persona_id'))).distinct()
                    for extension in personas:
                        ws.write('A%s' % i, str(extension.persona.identificacion()), formatoceldaleft)
                        ws.write('B%s' % i, str(extension.persona.nombre_completo_inverso_2()), formatoceldaleft)
                        ws.write('C%s' % i, str(extension.persona.mi_cargo_actual()), formatoceldaleft)
                        ws.write('D%s' % i, str(extension.persona.estado_civil()), formatoceldaleft)
                        ws.write('E%s' % i, str(extension.persona.sexo), formatoceldaleft)
                        ws.write('F%s' % i, str(extension.persona.mi_cargo_actual().regimenlaboral), formatoceldaleft)
                        ws.write('G%s' % i, str(extension.persona.mi_cargo_actual().unidadorganica), formatoceldaleft)
                        ws.write('H%s' % i, str(extension.get_tienehijos_display()), formatoceldaleft)
                        ws.write('I%s' % i, str(extension.hijos), formatoceldaleft)

                        i += 1
                    workbook.close()
                    output.seek(0)
                    filename = 'reporte_resumen_encuesta.xlsx'
                    response = HttpResponse(output,
                                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    response['Content-Disposition'] = 'attachment; filename=%s' % filename
                    return response
                except Exception as ex:
                    pass

            if action == 'addarchivoidioma':
                try:
                    puede_realizar_accion(request, 'sagest.puede_modificar_hoja_vida')
                    data['title'] = u'Adicionar archivo de idioma'
                    data['idioma'] = IdiomaDomina.objects.get(pk=int(request.GET['id']))
                    data['personaadmin'] = personaadmin = Persona.objects.get(pk=int(request.GET['ida']))
                    data['form'] = ArchivoIdiomaForm()
                    return render(request, "th_personal/addarchivoidioma.html", data)
                except:
                    pass

            if action == 'addidioma':
                try:
                    puede_realizar_accion(request, 'sagest.puede_modificar_hoja_vida')
                    data['title'] = u'Adicionar idioma'
                    data['personaadmin'] = personaadmin = Persona.objects.get(pk=int(request.GET['id']))
                    data['form'] = IdiomaDominaForm()
                    return render(request, "th_personal/addidioma.html", data)
                except:
                    pass

            if action == 'editidioma':
                try:
                    puede_realizar_accion(request, 'sagest.puede_modificar_hoja_vida')
                    data['title'] = u'Editar Idioma'
                    data['personaadmin'] = personaadmin = Persona.objects.get(pk=int(request.GET['ida']))
                    data['idioma'] = idioma = IdiomaDomina.objects.get(pk=request.GET['id'])
                    data['form'] = IdiomaDominaForm(initial={'idioma': idioma.idioma,
                                                             'escritura': idioma.escritura,
                                                             'lectura': idioma.lectura,
                                                             'oral': idioma.oral})
                    return render(request, "th_personal/editidioma.html", data)
                except Exception as ex:
                    pass

            if action == 'delidioma':
                try:
                    puede_realizar_accion(request, 'sagest.puede_modificar_hoja_vida')
                    data['title'] = u'Eliminar idioma'
                    data['personaadmin'] = personaadmin = Persona.objects.get(pk=int(request.GET['ida']))
                    data['idioma'] = idioma = IdiomaDomina.objects.get(pk=int(request.GET['id']))
                    return render(request, "th_personal/delidioma.html", data)
                except:
                    pass

            elif action == 'addarchivocontratofirmadodirector':
                try:
                    data['id'] = request.GET['id']
                    data['filtro'] = filtro = PersonaContratos.objects.get(pk=int(request.GET['id']))
                    form = ArchivoHistorialContratoDirectorForm()
                    data['action'] = 'addarchivocontratofirmadodirector'
                    data['form2'] = form
                    template = get_template("th_hojavida/modal/formarchivocontrato.html")
                    return JsonResponse({'result': True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(f'{ex}')
                    return HttpResponseRedirect(request.path)

            elif action == 'editarchivocontratofirmadodirector':
                try:
                    data['id'] = request.GET['id']
                    data['filtro'] = filtro = PersonaContratos.objects.get(pk=int(request.GET['id']))
                    form = ArchivoHistorialContratoForm()
                    data['action'] = 'editarchivocontratofirmadodirector'
                    data['form2'] = form
                    template = get_template("th_hojavida/modal/formarchivocontrato.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'historialarchivopersonal':
                try:
                    historial = None
                    if HistorialArchivosContratos.objects.filter(personacontrato=int(request.GET['id']),
                                                                 status=True).exists():
                        historial = HistorialArchivosContratos.objects.filter(personacontrato=int(request.GET['id']),
                                                                              status=True).order_by("id")
                    data['historialarchivo'] = historial
                    template = get_template("th_hojavida/historialcontratopersona.html")
                    return JsonResponse({'result': True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(f'{ex}')
                    return HttpResponseRedirect(request.path)

            elif action == 'pdf':
                try:
                    data['title'] = u'Adicionar Programa'
                    adduserdata(request, data)
                    persona = Persona.objects.get(pk=int(request.GET['id']))
                    # persona = request.session['persona']
                    profesor = persona.profesor().id
                    data['datospersona'] = persona
                    data['profesor'] = profesor
                    if ResponsableEvaluacion.objects.filter(status=True, activo=True).exists():
                        data['responsable'] = ResponsableEvaluacion.objects.filter(status=True, activo=True)[0]
                    data['periodoslect'] = MigracionEvaluacionDocente.objects.values('idperiodo','tipoeval').filter(idprofesor=profesor,idperiodo=request.GET['idperiodo']).distinct()
                    data['detalleevaluacion'] = migra = MigracionEvaluacionDocente.objects.filter(idprofesor=profesor,idperiodo=request.GET['idperiodo']).order_by('tipoeval','idperiodo','carrera','semestre','materia')
                    data['detalleevalconmodulo'] = migra.filter(modulo=1)
                    data['moduloevalcuatro'] = migra.filter(modulo=1,tipoeval=4)[0] if migra.filter(modulo=1,tipoeval=4).exists() else {}
                    data['detalleevalsinmodulo'] = migra.filter(modulo=0)
                    data['sinmoduloevalcuatro'] = migra.filter(modulo=0,tipoeval=4)[0] if migra.filter(modulo=0,tipoeval=4).exists() else {}
                    data['promperiodosinmodulo'] = promfinalc = round(null_to_numeric(migra.filter(modulo=0).aggregate(prom=Avg('promedioasignatura'))['prom']), 2)
                    data['promperiodoconmodulo'] = promfinal = round(null_to_numeric(migra.filter(modulo=1).aggregate(prom=Avg('promedioasignatura'))['prom']), 2)
                    if promfinal:
                        notaf = str(request.GET['idperiodo']) + "-" + persona.cedula + "-" + str(promfinal)
                    else:
                        notaf = str(request.GET['idperiodo']) + "-" + persona.cedula + "-" + str(promfinalc)
                    data['nomperiodo'] = request.GET['nomperiodo']
                    data['tipoev'] = request.GET['tipoev']
                    qrname = 'qrce_evam_' + request.GET['idperiodo'] + persona.cedula
                    folder = os.path.join(os.path.join(SITE_STORAGE, 'media', 'qrcode', 'evaluaciondocente', 'qr'))
                    # folder = '/mnt/nfs/home/storage/media/qrcode/evaluaciondocente/'
                    rutapdf = folder + qrname + '.pdf'
                    rutaimg = folder + qrname + '.png'
                    if os.path.isfile(rutapdf):
                        os.remove(rutaimg)
                        os.remove(rutapdf)
                    url = pyqrcode.create('http://sga.unemi.edu.ec//media/qrcode/evaluaciondocente/' + qrname + '.pdf')
                    imageqr = url.png(folder + qrname + '.png', 16, '#000000')
                    imagebarcode = code128.image(notaf).save(folder + qrname + "_bar.png")
                    data['qrname'] = 'qr' + qrname
                    data['fechactual'] = datetime.now().strftime("%d")  + '/' + datetime.now().strftime("%m") + '/' + datetime.now().strftime("%y")+ ' ' + datetime.now().strftime("%H:%M")
                    return conviert_html_to_pdfsave('pro_certificados/certificado_porperiodo.html',
                                                    {
                                                        'pagesize': 'A4',
                                                        'listadoevaluacion': data,
                                                    },qrname + '.pdf'
                                                    )
                except Exception as ex:
                    pass

            elif action == 'pdfmodelo2015':
                try:
                    data['title'] = u'Adicionar Programa'
                    adduserdata(request, data)
                    persona = Persona.objects.get(pk=int(request.GET['id']))
                    profesor = persona.profesor().id
                    data['datospersona'] = persona
                    data['profesor'] = profesor
                    if ResponsableEvaluacion.objects.filter(status=True, activo=True).exists():
                        data['responsable'] = ResponsableEvaluacion.objects.filter(status=True, activo=True)[0]
                    data['nomperiodo'] = request.GET['nomperiodo']
                    data['resultados'] = notaporcentaje = ResumenParcialEvaluacionIntegral.objects.filter(profesor=profesor,proceso=request.GET['idperiodo']).order_by('materia__asignaturamalla__malla__carrera__id', 'materia__asignaturamalla__nivelmalla__id')
                    data['porcentaje'] = round(null_to_numeric(notaporcentaje.aggregate(prom=Avg('totalmateriadocencia'))['prom']), 2)
                    data['fechactual'] = datetime.now().strftime("%d") + '/' + datetime.now().strftime("%m") + '/' + datetime.now().strftime("%y")+ ' ' + datetime.now().strftime("%H:%M")
                    notaporcen = "0"+ str(request.GET['idperiodo']) + "-" + persona.cedula + "-" + str(data['porcentaje'])
                    qrname = 'qrce_2015_0' + request.GET['idperiodo'] + persona.cedula
                    # folder = SITE_STORAGE + 'media/qrcode/evaluaciondocente/'
                    folder = os.path.join(os.path.join(SITE_STORAGE, 'media', 'qrcode', 'evaluaciondocente', 'qr'))
                    rutapdf = folder + qrname + '.pdf'
                    rutaimg = folder + qrname + '.png'
                    if os.path.isfile(rutapdf):
                        os.remove(rutaimg)
                        os.remove(rutapdf)
                    url = pyqrcode.create('http://sga.unemi.edu.ec//media/qrcode/evaluaciondocente/' + qrname + '.pdf')
                    imageqr = url.png(folder + qrname + '.png', 16, '#000000')
                    imagebarcode = code128.image(notaporcen).save(folder + qrname + "_bar.png")
                    data['qrname'] = 'qr' + qrname
                    return conviert_html_to_pdfsave(
                        'pro_certificados/pdfqrce_2015.html',
                        {
                            'pagesize': 'A4',
                            'listadoevaluacion': data,
                        },qrname + '.pdf'
                    )
                except Exception as ex:
                    pass

            elif action == 'cargasfamiliares':
                try:
                    data['title'] = u'Cargas familiares'
                    search = None
                    ids = None
                    periodosg=None
                    if 's' in request.GET:
                        search = request.GET['s']
                        ss = search.split(' ')
                        if len(ss) == 1:
                            periodosg = PeriodoGastosPersonales.objects.filter(Q(descripcion__icontains=search) |
                                                                               Q(anio__icontains=search) , status = True).distinct()
                    else:
                        periodosg = PeriodoGastosPersonales.objects.filter(status=True, anio__gte=2022).order_by('-anio')
                    lista = []
                    paging = MiPaginador(periodosg, 15)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['page'] = page
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['lisatdoperiodos'] = page.object_list
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    return render(request, "th_personal/viewperiodog.html", data)
                except Exception as ex:
                    pass

            elif action == 'decimo':
                try:
                    data['title'] = u'Configuración de Décimos'
                    search = None
                    ids = None
                    periodosg=None
                    periodosg = ConfiguraDecimo.objects.filter(status=True)
                    lista = []
                    paging = MiPaginador(periodosg, 15)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['page'] = page
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['registro'] = page.object_list
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    return render(request, "th_personal/viewdecimo.html", data)
                except Exception as ex:
                    pass

            elif action == 'importardecimo':
                try:
                    data['form'] = SagestImportarXLSForm()
                    data['action'] = 'importardecimo'
                    template = get_template("ajaxformmodal.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, 'mensaje': str(ex)}, safe=False)

            elif action == 'importarpersona':
                try:
                    data['form'] = SagestImportarXLSForm()
                    data['action'] = 'importarpersona'
                    data['id'] = encrypt_id(request.GET['id'])
                    template = get_template("th_personal/modal/formleyenda.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, 'mensaje': str(ex)}, safe=False)

            elif action == 'addperiodotthh':
                try:
                    form = PeriodoTTHHForm()
                    data['form'] = form
                    template = get_template('ajaxformmodal.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'editperiodotthh':
                try:
                    instancia = ConfiguraPeriodotthh.objects.get(id=encrypt_id(request.GET['id']))
                    form = PeriodoTTHHForm(initial=model_to_dict(instancia))
                    data['id'] = instancia.id
                    data['form'] = form
                    template = get_template('ajaxformmodal.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'periodorequisitotthh':
                try:
                    data['title'] = u'Configuración de periodos requisitos'
                    search = None
                    ids = None
                    periodosg=None
                    periodosg = ConfiguraPeriodotthh.objects.filter(status=True)
                    lista = []
                    paging = MiPaginador(periodosg, 15)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    ids_periodosg_pagina = [periodosg.id for periodosg in page.object_list]

                    subquery_requisitos = RequisitoPeriodotthh.objects.filter(periodotthh=OuterRef('pk'), status=True)
                    subquery_requisitos = subquery_requisitos.values('periodotthh').annotate(
                        numerorequisitos=Count('id')
                    )

                    periodosg_listado = ConfiguraPeriodotthh.objects.filter(id__in=ids_periodosg_pagina).\
                        annotate(tienepersonas=Exists(PersonaPeriodotthh.objects.filter(periodotthh_id=OuterRef('id'))),
                                 numeropersonas=Count('personaperiodotthh', filter=Q(personaperiodotthh__status=True)),
                                 numerorequisitos=Subquery(subquery_requisitos.values('numerorequisitos')))
                    data['page'] = page
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['registro'] = periodosg_listado
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    return render(request, "th_personal/viewperiodorequisitotthh.html", data)
                except Exception as ex:
                    pass

            elif action == 'listadorequisitos':
                try:
                    data['title'] = u'Requisitos'
                    data['periodotthh'] = periodotthh = ConfiguraPeriodotthh.objects.get(id=encrypt_id(request.GET['id']))
                    data['subtitle'] = u'Listado de requisitos solicitados ' + periodotthh.nombre
                    iter, search, url_vars, filtro = False, request.GET.get('s', ''), f'&action={action}', Q(status=True)
                    if search:
                        data['s'] = search
                        filtro = filtro & Q(nombre__unaccent__icontains=search)
                        url_vars += f'&s={search}'
                        iter = True
                    listarequisitos = periodotthh.requisitoperiodotthh_set.filter(filtro).order_by('-id')
                    paging = MiPaginador(listarequisitos, 10)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    ids_listarequisitos_pagina = [listarequisitos.id for listarequisitos in page.object_list]
                    listarequisitosperiodotthh = RequisitoPeriodotthh.objects.filter(id__in=ids_listarequisitos_pagina).\
                        annotate(enuso=Exists(DocumentoPersonaPeriodotthh.objects.filter(status=True,requisito_id=OuterRef('id'))))
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['iter'] = iter
                    data['url_vars'] = url_vars
                    data['listado'] = listarequisitosperiodotthh
                    return render(request, "th_personal/listadorequisitos.html", data)
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'personasperiodoth':
                try:
                    data['title'] = u'Personas'
                    data['periodotthh'] = periodotthh = ConfiguraPeriodotthh.objects.get(id=encrypt_id(request.GET['id']))
                    data['subtitle'] = u'Listado de personas ' + periodotthh.nombre
                    iter, search, url_vars, filtro = False, request.GET.get('s', ''), f'&action={action}', Q(status=True)
                    url_vars += f'&id=' + request.GET['id']
                    if search:
                        data['s'] = search
                        ss = search.split(' ')
                        if len(ss) == 1:
                            filtro = filtro & Q(Q(persona__nombres__icontains=search) |
                                                Q(persona__apellido1__icontains=search) |
                                                Q(persona__apellido2__icontains=search) |
                                                Q(persona__cedula__icontains=search) |
                                                Q(persona__pasaporte__icontains=search) |
                                                Q(persona__usuario__username__icontains=search))
                        else:
                            filtro = filtro & Q(Q(persona__apellido1__icontains=ss[0]) & Q(persona__apellido2__icontains=ss[1]))
                        url_vars += f'&s={search}'
                        iter = True
                    personasperiodotthh = periodotthh.personaperiodotthh_set.filter(filtro).order_by('persona__apellido1','persona__apellido2','persona__nombres')
                    paging = MiPaginador(personasperiodotthh, 10)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    ids_personasperiodotthh_pagina = [personasperiodotthh.id for personasperiodotthh in page.object_list]
                    personasperiodotthh = PersonaPeriodotthh.objects.filter(id__in=ids_personasperiodotthh_pagina).\
                        annotate(enuso=Exists(DocumentoPersonaPeriodotthh.objects.filter(status=True, personaperiodotthh_id=OuterRef('id'))))
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['iter'] = iter
                    data['url_vars'] = url_vars
                    data['estados_documentos'] = ESTADOS_DOCUMENTOS_REQUISITOS
                    data['listado'] = personasperiodotthh
                    return render(request, "th_personal/personasperiodoth.html", data)
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'validarrequisitotthh':
                try:
                    form = ValidarRequisitoTTHHForm()
                    data['form'] = form
                    data['personaperiodotthh'] = PersonaPeriodotthh.objects.get(pk=request.GET['id'])
                    template = get_template("th_personal/modal/formvalidarrequisito.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'addrequisitoperiodotthh':
                try:
                    form = RequisitoPeriodoTTHHForm()
                    data['form'] = form
                    data['id'] = encrypt_id(request.GET['id'])
                    data['switchery'] = True
                    template = get_template('ajaxformmodal.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'editrequisitoperiodotthh':
                try:
                    instancia = RequisitoPeriodotthh.objects.get(id=encrypt_id(request.GET['id']))
                    form = RequisitoPeriodoTTHHForm(initial=model_to_dict(instancia))
                    data['id'] = instancia.id
                    data['form'] = form
                    template = get_template('ajaxformmodal.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'descargarcomprimido':
                with transaction.atomic():
                    try:
                        ePersonPeriodo = PersonaPeriodotthh.objects.get(id=encrypt_id(request.GET['id']))
                        return descargar_comprimido_requisitos(request, ePersonPeriodo)
                    except Exception as ex:
                        messages.error(request, f"Error: {ex}")

            elif action == 'descargarcomprimidomasivo':
                with transaction.atomic():
                    try:
                        data['id'] = idperiodo = encrypt_id(request.GET['id'])
                        form = ComprimidoIngresoForm()
                        form.fields['requisito'].queryset = RequisitoPeriodotthh.objects.filter(status=True, periodotthh_id=idperiodo)
                        data['form']= form
                        template = get_template('th_personal/modal/formcomprimidomasivo.html')
                        return JsonResponse({'result':True, 'data':template.render(data)})
                    except Exception as ex:
                        return JsonResponse({'result': True, 'mensaje': f'Error: {ex}'})

            elif action == 'reporterequisitos':
                with transaction.atomic():
                    try:
                        idperiodo = encrypt_id(request.GET['id'])
                        estado, filtro=request.GET.get('estado_requisito',''), Q(status=True, periodotthh_id=idperiodo)
                        if estado:
                            filtro=filtro&Q(estado_requisito=int(estado))
                        ePersonasPeriodo = PersonaPeriodotthh.objects.filter(filtro)
                        wb = openxl.Workbook()
                        wb["Sheet"].title = "Reporte_requisitos"
                        ws = wb.active
                        style_title = openxlFont(name='Arial', size=16, bold=True)
                        style_cab = openxlFont(name='Arial', size=10, bold=True)
                        alinear = alin(horizontal="center", vertical="center")
                        response = HttpResponse(content_type="application/ms-excel")
                        response[
                            'Content-Disposition'] = 'attachment; filename=Reporte de requisitos de ingreso' + '-' + random.randint(
                            1, 10000).__str__() + '.xlsx'
                        ws.column_dimensions['B'].width = 35
                        ws.column_dimensions['C'].width = 15
                        ws.column_dimensions['D'].width = 20
                        ws.column_dimensions['E'].width = 25
                        ws.merge_cells('A1:J1')
                        ws['A1'] = 'REPORTE DE REQUISITOS DE INGRESO'
                        celda1 = ws['A1']
                        celda1.font = style_title
                        celda1.alignment = alinear

                        columns = [u"N°", u"Nombres",
                                   u"Cédula", u"Estado",
                                   u"Archivos Cargados","Pendientes", "Aprobados", "Corregir","Corregidos","Rechazados"]
                        row_num = 3
                        for col_num in range(0, len(columns)):
                            celda = ws.cell(row=row_num, column=(col_num + 1), value=columns[col_num])
                            celda.font = style_cab
                        row_num = 4
                        for idx, ePersonaPeriodo in enumerate(ePersonasPeriodo):
                            ws.cell(row=row_num, column=1, value=idx+1)
                            ws.cell(row=row_num, column=2, value=str(ePersonaPeriodo.persona))
                            ws.cell(row=row_num, column=3, value=str(ePersonaPeriodo.persona.cedula))
                            ws.cell(row=row_num, column=4, value=str(ePersonaPeriodo.get_estado_requisito_display()))
                            ws.cell(row=row_num, column=5, value=len(ePersonaPeriodo.documentos_subidos()))
                            totales = ePersonaPeriodo.totales()
                            ws.cell(row=row_num, column=6, value=totales['pendientes'])
                            ws.cell(row=row_num, column=7, value=totales['aprobados'])
                            ws.cell(row=row_num, column=8, value=totales['corregir'])
                            ws.cell(row=row_num, column=9, value=totales['corregidos'])
                            ws.cell(row=row_num, column=10, value=totales['rechazados'])
                            row_num += 1
                        wb.save(response)
                        return response
                    except Exception as ex:
                        messages.error(f'Error: {ex}')

            elif action == 'listadocontratosfirmados':
                try:
                    data['title'] = u'Listado de contratos pendientes por firmar director'
                    search = None
                    ids = None
                    if 's' in request.GET:
                        search = request.GET['s']
                        ss = search.split(' ')
                        if len(ss) == 1:
                            listadoprogramas = PersonaContratos.objects.filter(Q(persona__nombres__icontains=search) |
                                                                               Q(persona__apellido1__icontains=search) |
                                                                               Q(persona__apellido2__icontains=search) |
                                                                               Q(persona__cedula__icontains=search) |
                                                                               Q(persona__pasaporte__icontains=search) |
                                                                               Q(numerodocumento__icontains=search) |
                                                                               Q(persona__usuario__username__icontains=search), subio_archivo = True , status = True).distinct()
                        else:
                            listadoprogramas = PersonaContratos.objects.filter(Q(persona__apellido1__icontains=ss[0]) &
                                                                               Q(persona__apellido2__icontains=ss[1]), subio_archivo = True ,status = True).distinct()
                    else:
                        listadoprogramas = PersonaContratos.objects.filter(status=True , subio_archivo = True).order_by('persona__apellido1', 'persona__apellido2')

                    lista = []
                    for l in listadoprogramas:
                        if HistorialArchivosContratos.objects.filter(pk=l.ultimo_archivo).exists():
                            u = HistorialArchivosContratos.objects.get(pk=l.ultimo_archivo)
                            if u.estado_archivo == 2 or u.estado_archivo == 5:
                                lista.append(l)

                    paging = MiPaginador(lista, 15)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['page'] = page
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['listadoinscritos'] = page.object_list
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    return render(request, "th_personal/listadocontratosfirmados.html", data)
                except Exception as ex:
                    pass

            elif action == 'listadocontratosfirmadosfinalizados':
                try:
                    data['title'] = u'Listado de contratos firmados finalizados'
                    search = None
                    ids = None
                    if 's' in request.GET:
                        search = request.GET['s']
                        ss = search.split(' ')
                        if len(ss) == 1:
                            listadoprogramas = PersonaContratos.objects.filter(Q(persona__nombres__icontains=search) |
                                                                               Q(persona__apellido1__icontains=search) |
                                                                               Q(persona__apellido2__icontains=search) |
                                                                               Q(persona__cedula__icontains=search) |
                                                                               Q(persona__pasaporte__icontains=search) |
                                                                               Q(numerodocumento__icontains=search) |
                                                                               Q(persona__usuario__username__icontains=search), subio_archivo = True , status = True).distinct()
                        else:
                            listadoprogramas = PersonaContratos.objects.filter(Q(persona__apellido1__icontains=ss[0]) &
                                                                               Q(persona__apellido2__icontains=ss[1]), subio_archivo = True ,status = True).distinct()
                    else:
                        listadoprogramas = PersonaContratos.objects.filter(status=True , subio_archivo = True).order_by('persona__apellido1', 'persona__apellido2')

                    lista = []
                    for l in listadoprogramas:
                        if HistorialArchivosContratos.objects.filter(pk=l.ultimo_archivo).exists():
                            u = HistorialArchivosContratos.objects.get(pk=l.ultimo_archivo)
                            if u.estado_archivo == 3:
                                lista.append(l)

                    paging = MiPaginador(lista, 15)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['page'] = page
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['listadoinscritos'] = page.object_list
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    return render(request, "th_personal/listadocontratosfirmados_finalizados.html", data)
                except Exception as ex:
                    pass

            elif action == 'validarhorario':
                try:
                    fi, hi, hf = request.GET['fi'], request.GET['hi'], request.GET['hf']

                    mensaje = u"La %s ingresada es incorrecta. Por favor revisar."
                    if not fi:
                        return {"result": False, 'mensaje': mensaje % 'fecha'}

                    if not hi:
                        return {"result": False, 'mensaje': mensaje % 'hora de inicio'}

                    if not hf:
                        return {"result": False, 'mensaje': mensaje % u'hora de finalización'}

                    fecha = convertir_fecha_invertida(fi)
                    listaturnosocupados = validar_choque_horario_actividad_gestion(persona, fecha, hi, hf, request.session['periodo'])
                    if listaturnosocupados[0]:
                        data['choqueturno'] = listaturnosocupados[0]
                        data['listaturnosocupados'] = listaturnosocupados[1]
                        data['info'] = {'fecha':fecha, 'hi':hi, 'hf':hf}
                        template = get_template("th_hojavida/modal/listaturnosdisponibles.html")
                        return JsonResponse({"result": False, 'data': template.render(data)})
                    else:
                        return JsonResponse({"result": True, 'mensaje': 'Horario valido.'})

                except Exception as ex:
                    return JsonResponse({"result": False, 'mensaje': 'Error de conexión'})

            elif action == 'importarActividades':
                try:
                    data['id_persona'] = request.GET['id']
                    data['action'] = 'importarActividades'
                    template = get_template("th_personal/modal/importarActividades.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

            elif action == 'reporte_bitacora_excel':
                try:
                    fechadesde = request.GET['fecha_desde']
                    fechahasta = request.GET['fecha_hasta']
                    person = Persona.objects.get(pk=int(request.GET['ida']))
                    __author__ = 'Unemi'
                    title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                    title2 = easyxf( 'font: name Times New Roman, color-index black, bold on , height 250; alignment: horiz centre')
                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('Hoja1')
                    bitacoras = BitacoraActividadDiaria.objects.filter(status=True, fecha__date__gte=fechadesde,
                                                                       fecha__date__lte=fechahasta, persona=person)
                    ws.write_merge(0, 0, 0, 5, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                    ws.write_merge(1, 1, 0, 5, 'BITÁCORA DE ACTIVIDADES DIARIAS', title2)
                    ws.write_merge(2, 2, 0, 5, f'{person}', title2)
                    ws.write_merge(3, 3, 0, 5, f'DESDE  {fechadesde} HASTA {fechahasta}', title2)
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=bitacora' + random.randint(1, 10000).__str__() + '.xls'
                    columns = [
                        (u"DEPARTAMENTO", 6000),
                        (u"FECHA", 6000),
                        (u"TITULO", 6000),
                        (u"DESCRIPCIÓN", 15000),
                        (u"LINK", 6000),
                        (u"TIPO SISTEMA", 6000),
                    ]
                    row_num = 5
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], font_style)
                        ws.col(col_num).width = columns[col_num][1]

                    row_num = 6
                    for r in bitacoras:
                        campo1 = u"%s" % r.departamento.nombre if r.departamento else 'SIN DEPARTAMENTO'
                        campo2 = u"%s" % r.fecha
                        campo3 = u"%s" % r.titulo
                        campo4 = u"%s" % r.descripcion
                        campo5 = u"%s" % r.link
                        campo6 = u"%s" % r.get_tiposistema_display()
                        ws.write(row_num, 0, campo1, font_style2)
                        ws.write(row_num, 1, campo2, font_style2)
                        ws.write(row_num, 2, campo3, font_style2)
                        ws.write(row_num, 3, campo4, font_style2)
                        ws.write(row_num, 4, campo5, font_style2)
                        ws.write(row_num, 5, campo6, font_style2)
                        row_num += 1
                    wb.save(response)
                    return response
                except Exception as ex:
                    messages.error(request, str(ex))
                    #pass

            elif action == 'formatopersona':
                try:
                    __author__ = 'Unemi'
                    output = io.BytesIO()
                    workbook = xlsxwriter.Workbook(output)
                    ws = workbook.add_worksheet('formato')
                    ws.set_column(0, 36, 30)
                    formatoceldacenter = workbook.add_format({'border': 1, 'text_wrap': True, 'align': 'center', 'valign': 'vcenter'})
                    columns = [
                        (u"CEDULA", 4000),
                        (u"APELLIDOS Y NOMBRES", 8000),
                    ]
                    row_num = 0
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], formatoceldacenter)
                    row_num = 1
                    ws.write(row_num, 0, '0923354321', formatoceldacenter)
                    ws.write(row_num, 1, 'VITERI MARIDUEÑA ANGEL RODRIGO', formatoceldacenter)
                    row_num += 1
                    workbook.close()
                    output.seek(0)
                    filename = 'formato.xlsx'
                    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    response['Content-Disposition'] = 'attachment; filename=%s' % filename
                    return response
                except Exception as ex:
                    messages.error(request, str(ex))

            elif action == 'delbitacora':
                try:
                    data['title'] = u'Eliminar bitácora'
                    data['bitacora'] = BitacoraActividadDiaria.objects.get(pk=int(encrypt(request.GET['id'])))
                    data['idp'] = request.GET['idp']
                    return render(request, "th_personal/delbitacora.html", data)
                except Exception as ex:
                    pass

            elif action == 'informe-administrativo-posgrado':
                import unicodedata
                try:
                    data['title'] = u"INFORME DE ACTIVIDADES ADMINISTRATIVAS"
                    data['hoy'] = hoy = datetime.now()
                    data['per'] = per = Persona.objects.get(pk=int(request.GET['idp']))
                    if not request.GET.get('fi', '') or not request.GET.get('ff', ''):
                        return HttpResponseRedirect(f"/th_hojavida?info=Seleccione la fecha inicio/fin de la actividad.")

                    fi, ff = convertir_fecha_hora_invertida(request.GET['fi'] + " 00:01"), convertir_fecha_hora_invertida(request.GET['ff'] + " 23:59")
                    actividadesbitacora = BitacoraActividadDiaria.objects.filter(Q(fecha__gte=fi, fecha__lte=ff, persona=per, status=True)).order_by('fecha')
                    contrato = ContratoDip.objects.filter(
                        Q(persona=per, fechainicio__lte=ff, fechainicio__lt=ff, fechafin__gt=fi, fechafin__gte=fi,
                          status=True)).order_by('-fecha_creacion').first()
                    if contrato:
                        data['fi'], data['ff'] = fi, ff
                        data['responsable'] = u"%s" % contrato.seccion.responsable
                        data['contrato'] = contrato
                        data['actividades'] = actividadesbitacora
                        data['carreras'] = ContratoCarrera.objects.filter(contrato=contrato, status=True)
                        nombres = unicodedata.normalize('NFD', u"%s" % per.un_nombre_dos_apellidos()).encode('ascii', 'ignore').decode("utf-8").upper()
                        return conviert_html_to_pdf_name('th_personal/informe_actividad_mensual_administrativo_posgrado.html', {"data": data}, f"INFORME ADMINISTRATIVO - {nombres} {hoy.hour}{hoy.minute}{hoy.second}")
                    else:
                        return HttpResponseRedirect(f"/th_hojavida?info=Estimad{'o' if per.es_mujer() else 'a'} {per}, no se encontró un contrato vigente.")
                except Exception as ex:
                    return HttpResponseRedirect(f"/th_hojavida?info=Error de conexión. {ex.__str__()}")

            elif action == 'datossindistributivo':
                try:
                    __author__ = 'Unemi'
                    style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
                    style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
                    style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
                    title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                    style1 = easyxf(num_format_str='D-MMM-YY')
                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('exp_xls_post_part')
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=datos' + random.randint(1, 10000).__str__() + '.xls'
                    columns = [
                        (u"CEDULA", 4000),
                        (u"EMPLEADO", 8000),
                        (u"FECHA DE NACIMIENTO", 5000),
                        (u"CORREO INSTITUCIONAL", 2500),
                        (u"CORREO PERSONAL", 2500),
                        (u"CELULAR", 2500),
                        (u"TELEFONO", 2500),
                        (u"TELEFONO CONV", 2500),
                        (u"TELEFONO2", 2500),
                    ]
                    row_num = 0
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], font_style)
                        ws.col(col_num).width = columns[col_num][1]
                    date_format = xlwt.XFStyle()
                    date_format.num_format_str = 'yyyy-mm-dd'
                    date_formatreverse = xlwt.XFStyle()
                    date_formatreverse.num_format_str = 'dd/mm/yyyy'
                    personaladministrativo = Persona.objects.filter(Q(administrativo__isnull=False) | Q(profesor__isnull=False)).distinct()
                    personals = DistributivoPersona.objects.filter(status=True, estadopuesto__id=PUESTO_ACTIVO_ID).order_by('unidadorganica','persona__apellido1')
                    row_num = 1
                    i = 0
                    for persona in personals:
                        i += 1
                        ws.write(row_num, 0, persona.cedula, font_style2)
                        ws.write(row_num, 1, persona.apellido1 + ' ' + persona.apellido2 + ' ' + persona.nombres, font_style2)
                        ws.write(row_num, 2, persona.nacimiento, style1)
                        ws.write(row_num, 3, persona.emailinst, style1)
                        ws.write(row_num, 4, persona.email, style1)
                        ws.write(row_num, 5, persona.telefono, style1)
                        ws.write(row_num, 6, persona.telefono_conv, style1)
                        ws.write(row_num, 7, persona.telefono2, style1)
                        row_num += 1
                    wb.save(response)
                    return response
                except Exception as ex:
                    pass

            elif action == 'datos':
                try:
                    __author__ = 'Unemi'
                    style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
                    style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
                    style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
                    title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                    style1 = easyxf(num_format_str='D-MMM-YY')
                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('exp_xls_post_part')
                    # ws.write_merge(0, 0, 0, 10, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=datos' + random.randint(1, 10000).__str__() + '.xls'
                    columns = [
                        (u"CEDULA", 4000),
                        (u"EMPLEADO", 8000),
                        (u"FECHA DE NACIMIENTO", 5000),
                        (u"CORREO INSTITUCIONAL", 2500),
                        (u"CORREO PERSONAL", 2500),
                        (u"CELULAR", 2500),
                        (u"TELEFONO CONV", 2500),
                        (u"TELEFONO2", 2500),
                        (u"FECHA DE INGRESO", 2500),
                        (u"AÑOS EN LA INSTITUCIÓN ", 2500),
                        (u"CARGO", 4000),
                        (u"DEPARTAMENTO", 4000),
                        (u"REGIMEN", 4000),
                        (u"RMU PUESTO", 4000),
                        (u"RMU ESCALA", 4000),
                        (u"RMU SOBREVALORADO", 4000),
                    ]
                    row_num = 0
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], font_style)
                        ws.col(col_num).width = columns[col_num][1]
                    date_format = xlwt.XFStyle()
                    date_format.num_format_str = 'yyyy-mm-dd'
                    date_formatreverse = xlwt.XFStyle()
                    date_formatreverse.num_format_str = 'dd/mm/yyyy'
                    personaladministrativo = Persona.objects.filter(Q(administrativo__isnull=False) | Q(profesor__isnull=False)).distinct()
                    personals = DistributivoPersona.objects.filter(status=True, estadopuesto__id=PUESTO_ACTIVO_ID).order_by('unidadorganica','persona__apellido1')
                    row_num = 1
                    i = 0
                    for personal in personals:
                        i += 1
                        fecha_inicial = personal.persona.ingreso_personal_fecha(personal.regimenlaboral)
                        anios = 0
                        if fecha_inicial:
                            fecha_actual = datetime.now().date()
                            anios = (fecha_actual.year)-(fecha_inicial.year)
                        # ws.write(row_num, 0, i, font_style2)
                        ws.write(row_num, 0, personal.persona.cedula, font_style2)
                        ws.write(row_num, 1, personal.persona.apellido1 + ' ' + personal.persona.apellido2 + ' ' + personal.persona.nombres, font_style2)
                        ws.write(row_num, 2, personal.persona.nacimiento, style1)
                        ws.write(row_num, 3, personal.persona.emailinst, style1)
                        ws.write(row_num, 4, personal.persona.email, style1)
                        ws.write(row_num, 5, personal.persona.telefono, style1)
                        ws.write(row_num, 6, personal.persona.telefono_conv, style1)
                        ws.write(row_num, 7, personal.persona.telefono2, style1)
                        ws.write(row_num, 8, personal.persona.ingreso_personal_fecha(personal.regimenlaboral), style1)
                        ws.write(row_num, 9, str(anios), style1)
                        ws.write(row_num, 10, personal.denominacionpuesto.descripcion, font_style2)
                        ws.write(row_num, 11, personal.unidadorganica.nombre, date_format)
                        ws.write(row_num, 12, personal.regimenlaboral.descripcion, font_style2)
                        ws.write(row_num, 13, personal.rmupuesto, font_style2)
                        ws.write(row_num, 14, personal.rmuescala, font_style2)
                        ws.write(row_num, 15, personal.rmusobrevalorado, font_style2)
                        row_num += 1
                    wb.save(response)
                    return response
                except Exception as ex:
                    pass

            elif action =='repdecbienes':
                try:
                    __author__ = 'Unemi'

                    output = io.BytesIO()
                    workbook = xlsxwriter.Workbook(output)
                    ws = workbook.add_worksheet('diagnostico')
                    ws.set_column(0, 36, 30)

                    formatoceldaleft = workbook.add_format({'text_wrap': True, 'align': 'left'})

                    formatoceldacenter = workbook.add_format(
                        {'border': 1, 'text_wrap': True, 'align': 'center', 'valign': 'vcenter'})

                    formatoceldagris = workbook.add_format({'align': 'center', 'border': 1, 'text_wrap': True, 'fg_color': '#B6BFC0'})
                    # ws.write_merge(0, 0, 0, 10, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                    personaladministrativo = Persona.objects.filter(Q(administrativo__isnull=False) | Q(profesor__isnull=False)).distinct()
                    personals = DistributivoPersona.objects.filter(status=True).distinct('persona_id').order_by('persona_id')
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=declaracionbienes' + random.randint(1, 10000).__str__() + '.xls'
                    columns = [
                        (u"NOMBRE", 4000),
                        (u"TIPO DECLARACION", 8000),
                        (u"CODIGO DE BARRA", 5000),
                        (u"FECHA", 5000),
                    ]
                    row_num = 0
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], formatoceldacenter)

                    row_num = 1
                    for per in personals:
                        if DeclaracionBienes.objects.values('id').filter(status = True, persona=per.persona, verificado=True).exists():

                            dis = DeclaracionBienes.objects.filter(status = True, persona=per.persona, verificado=True)
                            if len(dis)>1:
                                val = len(dis)-1
                                ws.merge_range('A%s:A%s'%(row_num+1,row_num+1+val), str(per.persona), formatoceldacenter)
                                for dec in dis:
                                    if dec.tipodeclaracion:
                                        varaux = dec.get_tipodeclaracion_display()
                                    else:
                                        varaux = "-"
                                    ws.write(row_num, 1, str(varaux), formatoceldacenter)
                                    ws.write(row_num, 2, str(dec.codigobarra), formatoceldacenter)
                                    ws.write(row_num, 3, str(dec.fecha), formatoceldacenter)
                                    row_num +=1

                            else:
                                for dec in dis:
                                    if dec.tipodeclaracion:
                                        varaux = dec.get_tipodeclaracion_display()
                                    else:
                                        varaux = "-"
                                    ws.write(row_num, 0, str(per.persona), formatoceldacenter)
                                    ws.write(row_num, 1, str(varaux), formatoceldacenter)
                                    ws.write(row_num, 2, str(dec.codigobarra), formatoceldacenter)
                                    ws.write(row_num, 3, str(dec.fecha), formatoceldacenter)
                                    row_num +=1


                    workbook.close()
                    output.seek(0)
                    filename = 'declaracion_bienes.xlsx'
                    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    response['Content-Disposition'] = 'attachment; filename=%s' % filename
                    return response
                except Exception as ex:
                    pass

            elif action == 'detalledatoimp':
                try:
                    personaladministrativo = Persona.objects.filter(
                        Q(administrativo__isnull=False) | Q(profesor__isnull=False)).distinct()
                    listaadministrativos = personaladministrativo.filter(
                        Q(titulacion__verificado=False, titulacion__cursando=False, titulacion__archivo__isnull=False,
                          titulacion__registroarchivo__isnull=False) |
                        Q(capacitacion__verificado=False, capacitacion__archivo__isnull=False) |
                        Q(declaracionbienes__verificado=False, declaracionbienes__archivo__isnull=False) |
                        Q(cuentabancariapersona__verificado=False, cuentabancariapersona__archivo__isnull=False) |
                        Q(experiencialaboral__verificado=False, experiencialaboral__archivo__isnull=False)).distinct()
                    __author__ = 'Unemi'
                    style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
                    style_nb = easyxf('font: name Times New Roman, color-index blue, bold on',
                                      num_format_str='#,##0.00')
                    style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
                    title = easyxf(
                        'font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                    style1 = easyxf(num_format_str='D-MMM-YY')
                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('Listado')
                    # ws.write_merge(0, 0, 0, 10, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=Lista de nuevos datos ' + random.randint(1,
                                                                                                                     10000).__str__() + '.xls'
                    columns = [
                        (u"CEDULA", 4000),
                        (u"NOMBRE", 10000),
                        (u"TITULOS", 4000),
                        (u"CURSOS", 4000),
                        (u"DECLARACIONES", 4500),
                        (u"CTAS. BANCARIAS", 4500),
                        (u"EXPERIENCIAS", 4000),
                    ]

                    row_num = 0
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], font_style)
                        ws.col(col_num).width = columns[col_num][1]
                    date_format = xlwt.XFStyle()
                    date_format.num_format_str = 'yyyy-mm-dd'
                    date_formatreverse = xlwt.XFStyle()
                    date_formatreverse.num_format_str = 'dd/mm/yyyy'
                    row_num = 1
                    for personal in listaadministrativos:
                        campo0 = personal.cedula
                        campo1 = u'%s %s %s' % (personal.apellido1, personal.apellido2, personal.nombres)
                        campo2 = personal.cantidad_titulos_nuevos()
                        campo3 = personal.cantidad_cursos_nuevos()
                        campo4 = personal.cantidad_declaraciones_nuevas()
                        campo5 = personal.cantidad_cuentas_bancarias_nuevas()
                        campo6 = personal.cantidad_experiencia_laboral_nuevas()
                        ws.write(row_num, 0, campo0, font_style2)
                        ws.write(row_num, 1, campo1, font_style2)
                        ws.write(row_num, 2, campo2, font_style2)
                        ws.write(row_num, 3, campo3, font_style2)
                        ws.write(row_num, 4, campo4, font_style2)
                        ws.write(row_num, 5, campo5, font_style2)
                        ws.write(row_num, 6, campo6, font_style2)
                        row_num += 1
                    wb.save(response)
                    return response
                except Exception as ex:
                    pass

            elif action == 'rephojavida':
                try:
                    data['fechaactual'] = datetime.now().date()
                    data['title'] = u'Reporte de Actualización de Hoja de Vida Institucional de los Servidores Publicos'
                    data['regimenlaborales'] = regimenslaborales = RegimenLaboral.objects.filter(status=True, pk__in=[1,2,4])
                    regimenlaboral, desde, hasta, search, filtrosupdate, filtros,  filtrodistributivo, url_vars = request.GET.get('regimenlaboral', ''),  request.GET.get('desde', ''),  request.GET.get('hasta', ''),  request.GET.get('search', ''), Q(status=True), Q(status=True),  Q(status=True), ''
                    if desde:
                        data['desde'] = desdestr = desde
                        url_vars += "&desde={}".format(desde)
                        filtros = filtros & Q(fecha_creacion__gte=desde)
                        filtrosupdate = filtrosupdate & Q(fecha_modificacion__gte=desde)
                    else:
                        rangodias = calendar.monthrange(datetime.now().date().year, datetime.now().date().month)
                        desde = convertir_fecha_invertida('{}-{}-{}'.format(datetime.now().date().year, datetime.now().date().month, rangodias[0]))
                        data['desde'] = desdestr = str(desde)
                        url_vars += "&desde={}".format(desde)
                        filtros = filtros & Q(fecha_creacion__gte=desde)
                        filtrosupdate = filtrosupdate & Q(fecha_modificacion__gte=desde)
                    if hasta:
                        data['hasta'] = hastastr = hasta
                        url_vars += "&hasta={}".format(hasta)
                        filtros = filtros & Q(fecha_creacion__lte=hasta)
                        filtrosupdate = filtrosupdate & Q(fecha_modificacion__lte=hasta)
                    else:
                        rangodias = calendar.monthrange(datetime.now().date().year, datetime.now().date().month)
                        hasta = convertir_fecha_invertida('{}-{}-{}'.format(datetime.now().date().year, datetime.now().date().month, rangodias[1]))
                        data['hasta'] = hastastr = str(hasta)
                        url_vars += "&hasta={}".format(hasta)
                        filtros = filtros & Q(fecha_creacion__lte=hasta)
                        filtrosupdate = filtrosupdate & Q(fecha_modificacion__lte=hasta)
                    if search:
                        data['search'] = search
                        filtrodistributivo = filtrodistributivo & (Q(persona__cedula__icontains=search) | Q(persona__apellido1__icontains=search))
                        url_vars += '&search={}'.format(search)
                    if regimenlaboral:
                        data['regimenlaboral'] = int(regimenlaboral)
                        filtrodistributivo = filtrodistributivo & (Q(regimenlaboral_id=int(regimenlaboral)))
                        url_vars += '&regimenlaboral={}'.format(regimenlaboral)
                    else:
                        filtrodistributivo = filtrodistributivo & (Q(regimenlaboral__in=[1,2,4]))
                    url_vars += '&action={}'.format(action)
                    data["url_vars"] = url_vars
                    qscertificados = CertificacionPersona.objects.filter(filtros)
                    qscertificadosupdate = CertificacionPersona.objects.filter(filtrosupdate)
                    qstitulacion = Titulacion.objects.filter(filtros)
                    qstitulacionupdate = Titulacion.objects.filter(filtrosupdate)
                    qsdeclaracion = DeclaracionBienes.objects.filter(filtros)
                    qsdeclaracionupdate = DeclaracionBienes.objects.filter(filtrosupdate)
                    qscuentabancaria = CuentaBancariaPersona.objects.filter(filtros)
                    qscuentabancariaupdate = CuentaBancariaPersona.objects.filter(filtrosupdate)
                    qsexperiencia = ExperienciaLaboral.objects.filter(filtros)
                    qsexperienciaupdate = ExperienciaLaboral.objects.filter(filtrosupdate)
                    personasconcreacion = list(qscertificados.values_list('persona_id', flat=True)) + list(qstitulacion.values_list('persona_id', flat=True)) + list(qsdeclaracion.values_list('persona_id', flat=True)) + list(qscuentabancaria.values_list('persona_id', flat=True)) + list(qsexperiencia.values_list('persona_id', flat=True))
                    personasconactualizacion = list(qscertificadosupdate.values_list('persona_id', flat=True)) + list(qstitulacionupdate.values_list('persona_id', flat=True)) + list(qsdeclaracionupdate.values_list('persona_id', flat=True)) + list(qscuentabancariaupdate.values_list('persona_id', flat=True)) + list(qsexperienciaupdate.values_list('persona_id', flat=True))
                    personasact = list(dict.fromkeys(personasconcreacion + personasconactualizacion))
                    querytot = DistributivoPersona.objects.filter(estadopuesto_id=1).filter(Q(persona__administrativo__isnull=False) | Q(persona__profesor__isnull=False)).filter(filtrodistributivo).order_by('regimenlaboral','persona__apellido1')
                    data['personasactualizadas'] = query = DistributivoPersona.objects.filter(estadopuesto_id=1).filter(Q(persona__administrativo__isnull=False) | Q(persona__profesor__isnull=False)).filter(filtrodistributivo).filter(persona__in=personasact).order_by('regimenlaboral','persona__apellido1')
                    querynoact = DistributivoPersona.objects.filter(estadopuesto_id=1).filter(Q(persona__administrativo__isnull=False) | Q(persona__profesor__isnull=False)).filter(filtrodistributivo).exclude(persona__in=personasact).order_by('regimenlaboral','persona__apellido1')
                    data['listcount'] = totpersonasact = query.count()
                    data['listcountstr'] = str(totpersonasact)
                    data['listcountquerynoact'] = totpersonasnoact = querynoact.count()
                    data['listcounttot'] = totpersonas = querytot.count()
                    data['porceactualizado'] = porceactualizado = round(null_to_decimal((totpersonasact*100)/totpersonas),2)
                    data['porcenoactualizado'] = porcenoactualizado = round(null_to_decimal((totpersonasnoact*100)/totpersonas),2)
                    listatotalesregimenlaborales = []
                    listanototalesregimenlaborales = []
                    for rl in regimenslaborales:
                        tot = query.filter(regimenlaboral=rl).count()
                        listatotalesregimenlaborales.append({'pk': rl.pk, 'query': query.filter(regimenlaboral=rl), 'desc': rl.descripcion, 'tot': tot, 'porc': round(null_to_decimal((tot*100)/totpersonas),2)})
                    data['listatotalesregimenlaborales'] = listatotalesregimenlaborales
                    for rl in regimenslaborales:
                        tot = querynoact.filter(regimenlaboral=rl).count()
                        listanototalesregimenlaborales.append({'pk': rl.pk, 'query': querynoact.filter(regimenlaboral=rl), 'desc': rl.descripcion, 'tot': tot, 'porc': round(null_to_decimal((tot*100)/totpersonas),2)})
                    data['listanototalesregimenlaborales'] = listanototalesregimenlaborales

                    if 'pdf' in request.GET:
                        return conviert_html_to_pdf(
                            'th_personal/pdfactualizacionhv.html',
                            {
                                'pagesize': 'A4',
                                'data': data,
                            }
                        )

                    if 'pdfdet' in request.GET:
                        return conviert_html_to_pdf(
                            'th_personal/pdfactualizacionhvdet.html',
                            {
                                'pagesize': 'A4',
                                'data': data,
                            }
                        )

                    if 'excel' in request.GET:
                        __author__ = 'Unemi'
                        borders = Borders()
                        borders.left = Borders.THIN
                        borders.right = Borders.THIN
                        borders.top = Borders.THIN
                        borders.bottom = Borders.THIN
                        align = Alignment()
                        font_style = XFStyle()
                        font_style.font.bold = True
                        font_style.borders = borders
                        font_style.alignment = align
                        font_style2 = XFStyle()
                        font_style2.font.bold = False
                        font_style2.borders = borders
                        font_style2.alignment = align
                        style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
                        style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
                        style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
                        title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                        style1 = easyxf(num_format_str='D-MMM-YY')
                        title = easyxf('font: name Calibri, color-index black, bold on , height 260; alignment: horiz centre')
                        fuentecabecera = easyxf('font: name Calibri, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: horiz centre; borders: left thin, right thin, top thin, bottom thin')
                        fuentenormal = easyxf('font: name Calibri, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
                        fuentetexto = easyxf('font: name Calibri, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
                        font_style = XFStyle()
                        font_style.font.bold = True
                        font_style2 = XFStyle()
                        font_style2.font.bold = False
                        wb = Workbook(encoding='utf-8')
                        ws = wb.add_sheet('consolidado')
                        # ws.write_merge(0, 0, 0, 10, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                        response = HttpResponse(content_type="application/ms-excel")
                        response['Content-Disposition'] = 'attachment; filename=reporte_actualizacionhv_' + random.randint(1, 10000).__str__() + '.xls'
                        ws.write_merge(0, 3, 0, 3, 'REPORTE DE ACTUALIZACIÓN DE HOJA DE VIDA INSTITUCIONAL DE LOS SERVIDORES PÚBLICOS DE UNEMI', title)
                        ws.write_merge(4, 4, 0, 3, '{} - {}'.format(desdestr, hastastr), fuentenormal)
                        columns = [
                            (u"BANDERA ACTUALIZACIÓN", 8000),
                            (u"RÉGIMEN LABORAL", 10000),
                            (u"ACTUALIZACIÓN DE DATOS", 10000),
                            (u"VALOR PORCENTUAL", 4000),
                        ]
                        row_num = 5
                        for col_num in range(len(columns)):
                            ws.write(row_num, col_num, columns[col_num][0], fuentecabecera)
                            ws.col(col_num).width = columns[col_num][1]
                        row_num = 6
                        ws.write_merge(row_num, row_num, 0, 0, 'ACTUALIZADO', font_style2)
                        for q in listatotalesregimenlaborales:
                            ws.write_merge(row_num, row_num, 1, 1, q['desc'], font_style2)
                            for qs in q['query']:
                                ws.write_merge(row_num, row_num, 2, 2, qs.persona.__str__(), font_style2)
                                ws.write_merge(row_num, row_num, 3, 3, '', font_style2)
                                row_num += 1
                            if not q['query']:
                                row_num += 1
                            ws.write_merge(row_num, row_num, 1, 2, 'TOTAL {}'.format(q['desc']), font_style2)
                            ws.write_merge(row_num, row_num, 3, 3, q['porc'], font_style2)
                            row_num += 1
                        ws.write_merge(row_num, row_num, 0, 2, 'TOTAL ACTUALIZADO', font_style2)
                        ws.write_merge(row_num, row_num, 3, 3, porceactualizado, font_style2)
                        row_num += 1
                        ws.write_merge(row_num, row_num, 0, 0, 'NO ACTUALIZADO', font_style2)
                        for q in listanototalesregimenlaborales:
                            ws.write_merge(row_num, row_num, 1, 1, q['desc'], font_style2)
                            for qs in q['query']:
                                ws.write_merge(row_num, row_num, 2, 2, qs.persona.__str__(), font_style2)
                                ws.write_merge(row_num, row_num, 3, 3, '', font_style2)
                                row_num += 1
                            if not q['query']:
                                row_num += 1
                            ws.write_merge(row_num, row_num, 1, 2, 'TOTAL {}'.format(q['desc']), font_style2)
                            ws.write_merge(row_num, row_num, 3, 3, q['porc'], font_style2)
                            row_num += 1
                        ws.write_merge(row_num, row_num, 0, 2, 'TOTAL NO ACTUALIZADO', font_style2)
                        ws.write_merge(row_num, row_num, 3, 3, porcenoactualizado, font_style2)
                        row_num += 1
                        ws.write_merge(row_num, row_num, 0, 2, 'TOTAL GENERAL', font_style2)
                        ws.write_merge(row_num, row_num, 3, 3, 100, font_style2)
                        date_format = xlwt.XFStyle()
                        date_format.num_format_str = 'yyyy-mm-dd'
                        date_formatreverse = xlwt.XFStyle()
                        date_formatreverse.num_format_str = 'dd/mm/yyyy'
                        wb.save(response)
                        return response

                    paging = MiPaginador(query, 25)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['listado'] = page.object_list
                    return render(request, "th_personal/viewactualizacionhv.html", data)
                except Exception as ex:
                    print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
                    messages.error(request, ex)

            elif action == 'repdecimo':
                try:
                    distributivo = DistributivoPersona.objects.values_list('persona_id',flat=True).filter(status=True)
                    personasdecimo = RegistroDecimo.objects.filter(activo=True, status=True, estado=2, persona__id__in=distributivo)
                    __author__ = 'Unemi'
                    borders = Borders()
                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style.borders = borders
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    font_style2.borders = borders
                    title = easyxf('font: name Calibri, color-index black, bold on , height 260; alignment: horiz centre')
                    fuentecabecera = easyxf('font: name Calibri, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: horiz centre; borders: left thin, right thin, top thin, bottom thin')
                    fuentenormal = easyxf('font: name Calibri, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
                    fuentetexto = easyxf('font: name Calibri, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('consolidado')
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=reporte_decimo_' + random.randint(1, 10000).__str__() + '.xls'
                    ws.write_merge(0, 0, 0, 3, 'REPORTE DE ACTUALIZACIÓN DE DÉCIMO', title)
                    columns = [
                        (u"CEDULA", 8000),
                        (u"PASAPORTE", 8000),
                        (u"TRABAJADOR", 10000),
                        (u"CARGO", 10000),
                        (u"UNIDAD ORGÁNICA", 10000),
                        (u"SELECCIÓN", 10000),
                    ]
                    row_num = 2
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], fuentecabecera)
                        ws.col(col_num).width = columns[col_num][1]
                    row_num = 3
                    for decimo in personasdecimo:
                        cargo = decimo.persona.mi_cargo_actualadm()
                        if not cargo:
                            cargo = decimo.persona.mi_cargo_actual_docente()
                        ws.write(row_num, 0, str(decimo.persona.cedula), font_style2)
                        ws.write(row_num, 1, str(decimo.persona.pasaporte), font_style2)
                        ws.write(row_num, 2, str(decimo.persona.nombre_completo()), font_style2)
                        if cargo:
                            ws.write(row_num, 3, str(cargo.denominacionpuesto), font_style2)
                            ws.write(row_num, 4, str(cargo.unidadorganica), font_style2)
                        ws.write(row_num, 5, str(decimo.get_seleccion_display()), font_style2)
                        row_num+=1

                    wb.save(response)
                    return response
                except Exception as ex:
                    print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
                    messages.error(request, ex)

            elif action == 'consultatotalcontratos':
                numero = 0
                fecha_inicio = datetime.strptime(request.GET['finicio'], '%d-%m-%Y')
                fecha_fin = datetime.strptime(request.GET['ffin'], '%d-%m-%Y')
                lista = []
                contratos = PersonaContratos.objects.filter(
                    Q(status=True) & Q(fecha_creacion__gte=fecha_inicio) & Q(fecha_creacion__lte=fecha_fin)).order_by(
                    'id')
                for con in contratos:
                    if HistorialArchivosContratos.objects.filter(pk=con.ultimo_archivo).exists():
                        c = HistorialArchivosContratos.objects.get(pk=con.ultimo_archivo)
                        if c.estado_archivo == 2:
                            lista.append(con)
                numero = len(lista)
                response = JsonResponse({'state': True, 'totinformes': numero, })
                return HttpResponse(response.content)

            elif action == 'bajarconsolidadocontratoszip':
                try:
                    fecha_inicio = datetime.strptime(request.GET['finicio'], '%d-%m-%Y').date()
                    fecha_fin = datetime.strptime(request.GET['ffin'], '%d-%m-%Y').date()
                    mes = fecha_inicio.month
                    anio = fecha_inicio.year
                    lista = []
                    contratos = PersonaContratos.objects.filter(
                        Q(status=True) & Q(fecha_creacion__gte=fecha_inicio) & Q(fecha_creacion__lte=fecha_fin)).order_by(
                        'id')
                    for con in contratos:
                        if HistorialArchivosContratos.objects.filter(pk=con.ultimo_archivo).exists():
                            c = HistorialArchivosContratos.objects.get(pk=con.ultimo_archivo)
                            if c.estado_archivo == 2:
                                lista.append(c)
                    dominiosistema = request.build_absolute_uri('/')[:-1].strip("/")

                    archivos_lista = []

                    directory = os.path.join(SITE_STORAGE, 'zipav')
                    try:
                        os.stat(directory)
                    except:
                        os.mkdir(directory)

                    url = os.path.join(SITE_STORAGE, 'media', 'zipav',
                                       'contratosconsolidados_{}_{}_{}.zip'.format(mes, anio,
                                                                                   random.randint(1, 10000).__str__()))
                    url_zip = url
                    fantasy_zip = zipfile.ZipFile(url, 'w')
                    nombredocente = ''
                    for inf in lista:
                        if inf.archivo:
                            nombre = remover_caracteres_especiales_unicode(
                                inf.personacontrato.persona.__str__().lower().replace(' ', '_')).lower().replace(' ', '_')
                            fantasy_zip.write(inf.archivo.path,
                                              '{}_{}_{}.pdf'.format(nombre, mes, anio))
                    fantasy_zip.close()
                    response = HttpResponse(open(url_zip, 'rb'), content_type='application/zip')
                    response['Content-Disposition'] = 'attachment; filename=contratosconsolidados_{}_{}_{}.zip'.format(mes, anio,
                                                                                                                       random.randint(
                                                                                                                           1,
                                                                                                                           10000).__str__())
                    return response
                except Exception as ex:
                    print(ex)
                    print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
                    messages.error(request, nombredocente)
                    return redirect('{}?action=viewdocentes&info={}&linea={}'.format(request.path, ex,
                                                                                     sys.exc_info()[-1].tb_lineno))

            elif action == 'bajarconsolidadocontratoszipsinfechas':
                try:

                    lista = []
                    contratos = PersonaContratos.objects.filter(status=True , subio_archivo = True).order_by('persona__apellido1', 'persona__apellido2')
                    for con in contratos:
                        if HistorialArchivosContratos.objects.filter(pk=con.ultimo_archivo).exists():
                            c = HistorialArchivosContratos.objects.get(pk=con.ultimo_archivo)
                            if c.estado_archivo == 2:
                                lista.append(c)
                    directory = os.path.join(SITE_STORAGE, 'zipav')
                    try:
                        os.stat(directory)
                    except:
                        os.mkdir(directory)

                    url = os.path.join(SITE_STORAGE, 'media', 'zipav',
                                       'contratosconsolidados_{}.zip'.format(random.randint(1, 10000).__str__()))
                    url_zip = url
                    fantasy_zip = zipfile.ZipFile(url, 'w')
                    nombredocente = ''
                    for inf in lista:
                        if inf.archivo:
                            nombre = remover_caracteres_especiales_unicode(
                                inf.personacontrato.persona.__str__().lower().replace(' ', '_')).lower().replace(' ', '_')
                            fantasy_zip.write(inf.archivo.path,
                                              '{}_{}.pdf'.format(nombre, inf.personacontrato.numerodocumento))
                    fantasy_zip.close()
                    response = HttpResponse(open(url_zip, 'rb'), content_type='application/zip')
                    response['Content-Disposition'] = 'attachment; filename=contratosconsolidados_{}.zip'.format(len(lista))
                    log(u'Descargo el consolidado de contratos personal firmados: %s' % persona, request, "edit")

                    return response
                except Exception as ex:
                    print(ex)
                    print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
                    messages.error(request, nombredocente)
                    return redirect('{}?action=viewdocentes&info={}&linea={}'.format(request.path, ex,
                                                                                     sys.exc_info()[-1].tb_lineno))

            elif action == 'bajarconsolidadocontratoszipfinalizados':
                try:

                    lista = []
                    contratos = PersonaContratos.objects.filter(status=True , subio_archivo = True).order_by('persona__apellido1', 'persona__apellido2')
                    for con in contratos:
                        if HistorialArchivosContratos.objects.filter(pk=con.ultimo_archivo).exists():
                            c = HistorialArchivosContratos.objects.get(pk=con.ultimo_archivo)
                            if c.estado_archivo == 3:
                                lista.append(c)
                    directory = os.path.join(SITE_STORAGE, 'zipav')
                    try:
                        os.stat(directory)
                    except:
                        os.mkdir(directory)

                    url = os.path.join(SITE_STORAGE, 'media', 'zipav',
                                       'contratosconsolidadosfinalizados_{}.zip'.format(random.randint(1, 10000).__str__()))
                    url_zip = url
                    fantasy_zip = zipfile.ZipFile(url, 'w')
                    nombredocente = ''
                    for inf in lista:
                        if inf.archivo:
                            nombre = remover_caracteres_especiales_unicode(
                                inf.personacontrato.persona.__str__().lower().replace(' ', '_')).lower().replace(' ', '_')
                            fantasy_zip.write(inf.archivo.path,
                                              '{}_{}.pdf'.format(nombre, inf.personacontrato.numerodocumento))
                    fantasy_zip.close()
                    response = HttpResponse(open(url_zip, 'rb'), content_type='application/zip')
                    response['Content-Disposition'] = 'attachment; filename=contratosconsolidadosfinalizados_{}.zip'.format(len(lista))
                    log(u'Descargo el consolidado de contratos personal firmados: %s' % persona, request, "edit")

                    return response
                except Exception as ex:
                    print(ex)
                    print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
                    messages.error(request, nombredocente)
                    return redirect('{}?action=viewdocentes&info={}&linea={}'.format(request.path, ex,
                                                                                     sys.exc_info()[-1].tb_lineno))

            elif action == 'descargajubilado':
                try:
                    # periodo = PeriodoPlanificacionTH.objects.get(pk=request.GET['id'])
                    __author__ = 'Unemi'

                    output = io.BytesIO()
                    workbook = xlsxwriter.Workbook(output)
                    ws = workbook.add_worksheet('diagnostico')
                    ws.set_column(0, 36, 30)

                    formatoceldaleft = workbook.add_format({'text_wrap': True, 'align': 'left'})

                    formatoceldacenter = workbook.add_format(
                        {'border': 1, 'text_wrap': True, 'align': 'center', 'valign': 'vcenter'})

                    formatoceldagris = workbook.add_format( {'align': 'center', 'border': 1, 'text_wrap': True, 'fg_color': '#B6BFC0'})

                    ws.write('A1', 'REPORTE DE DIAGNÓSTICO INSTITUCIONAL',formatoceldaleft)
                    ws.write('A2', 'FECHA %s'%(datetime.today().date()),formatoceldaleft)
                    ws.write('A4', 'IDENTIFICACIÓN',formatoceldagris)
                    ws.write('B4', 'NOMBRES',formatoceldagris)
                    ws.write('C4', 'FECHA DE NACIMIENTO',formatoceldagris)
                    ws.write('D4', 'FECHA DE INGRESO A LA INSTITUCIÓN',formatoceldagris)
                    ws.write('E4', 'RÉGIMEN LABORAL',formatoceldagris)
                    ws.write('F4', 'MODALIDAD LABORAL',formatoceldagris)
                    ws.write('G4', 'DENOMINACIÓN PUESTO',formatoceldagris)
                    ws.write('H4', 'UNIDAD ORGÁNICA',formatoceldagris)

                    plantillas = DistributivoPersonaHistorial.objects.filter(status=True,regimenlaboral_id=3).distinct('persona__id').order_by('persona__id')
                    i = 5
                    for  plantilla in plantillas:

                        ws.write('A%s'%i, plantilla.persona.identificacion(),formatoceldaleft)
                        ws.write('B%s'%i, str(plantilla.persona),formatoceldaleft)
                        ws.write('C%s'%i, plantilla.persona.nacimiento,formatoceldaleft)
                        ws.write('D%s'%i, plantilla.persona.fechaingresoies,formatoceldaleft)
                        ws.write('E%s'%i, str(plantilla.regimenlaboral),formatoceldaleft)
                        ws.write('F%s'%i, str(plantilla.modalidadlaboral),formatoceldaleft)
                        ws.write('G%s'%i, str(plantilla.denominacionpuesto),formatoceldaleft)
                        ws.write('H%s'%i, str(plantilla.unidadorganica),formatoceldaleft)
                        i+=1

                    workbook.close()
                    output.seek(0)
                    filename = 'reporte_jubilados.xlsx'
                    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    response['Content-Disposition'] = 'attachment; filename=%s' % filename
                    return response
                except Exception as ex:
                    pass

            elif action == 'delcertificado':
                try:
                    data['title'] = u'Eliminar Certificación'
                    data['certificado'] = CertificacionPersona.objects.get(pk=request.GET['id'])
                    return render(request, "th_personal/delcertificado.html", data)
                except Exception as ex:
                    pass

            elif action == 'adddocumentofirmado':
                try:
                    data['filtro'] = filtro = PersonaAcciones.objects.get(pk=int(request.GET['id']))
                    data['formmodal'] = AccionPersonalDocumentoForm()
                    template = get_template("th_hojavida/modal/formdocumentofirmado.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'historialfirmados':
                try:
                    data['filtro'] = filtro = PersonaAcciones.objects.get(pk=int(request.GET['id']))
                    template = get_template("th_hojavida/modal/historialdocumentofirmado.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as e:
                    print(e)
                    pass

            elif action == 'descargarcomprimido':
                try:
                    data['form'] = DescargarCompromidoForm()
                    template = get_template("th_hojavida/modal/formcomprimido.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'repcargas':
                try:
                    data['form'] = DescagarCargasFechasForm()
                    data['switchery']=True
                    data['id']=request.GET['id']
                    template = get_template("th_hojavida/modal/formrepcarga.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'familiares':
                try:
                    data['title'] = u'Listado de familiares registrados'
                    search = None
                    ids = None
                    search, filtro, url_vars = request.GET.get('s', ''), \
                                               (Q(status=True)), '&action=familiares'
                    if search:
                        data['s'] = search = request.GET['s'].strip()
                        ss = search.split(' ')
                        url_vars += "&s={}".format(search)
                        if len(ss) == 1:
                            filtro = filtro & (Q(nombres__icontains=search) |
                                               Q(apellido1__icontains=search) |
                                               Q(apellido2__icontains=search) |
                                               Q(cedula__icontains=search) |
                                               Q(pasaporte__icontains=search))
                        else:
                            filtro = filtro & (Q(apellido1__icontains=ss[0]) & Q(apellido2__icontains=ss[1]))
                    elif 'id' in request.GET:
                        ids = request.GET['id']
                        filtro = filtro & (Q(id=ids))

                    distributivo = DistributivoPersona.objects.values_list('persona_id')
                    listado = Persona.objects.filter(filtro,id__in=(distributivo))
                    encuesta = PersonaExtension.objects.filter(persona_id__in=(distributivo)).distinct()
                    data['total'] = total = listado.count()
                    data['total_hijos'] =  encuesta.filter(tienehijos=2).count()
                    data['total_no_hijos'] = encuesta.filter(tienehijos=3).count()
                    data['total_pendiente'] = encuesta.filter(tienehijos=1).count()
                    paging = MiPaginador(listado, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['page'] = page
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['listado'] = page.object_list
                    data['search'] = search if search else ""
                    data['url_vars'] = url_vars
                    data['ids'] = ids if ids else ""
                    return render(request, "th_personal/familiares/familiares.html", data)
                except Exception as ex:
                    pass

            elif action == 'detallepersonal':
                try:

                    data['bitacoras'] = bitacoras = BitacoraActividadDiaria.objects.filter(persona_id=int(request.GET['ida']), status=True).order_by('-fecha')
                    data['bitacora_anios'] = bitacoras.dates('fecha', 'year').distinct()
                    data['id_per_bitacora'] = int(request.GET['ida'])
                    data['per_bitacora'] = Persona.objects.get(pk=int(request.GET['ida']))
                    searchb = None
                    if 'sb' in request.GET:
                        if request.GET['sb']:
                            searchb = request.GET['sb'].strip()
                            bitacoras = bitacoras.filter(Q(titulo__icontains=searchb) |
                                                         Q(departamento__nombre__icontains=searchb) |
                                                         Q(descripcion__icontains=searchb) |
                                                         Q(link__icontains=searchb))

                    aniobit = None
                    if 'aniobit' in request.GET:
                        if request.GET['aniobit']:
                            aniobit = request.GET['aniobit'].strip()
                            bitacoras = bitacoras.filter(fecha__year=aniobit)

                    mesbit = None
                    if 'mesbit' in request.GET:
                        if request.GET['mesbit']:
                            mesbit = request.GET['mesbit'].strip()
                            bitacoras = bitacoras.filter(fecha__month=mesbit)

                    pagingbitacora = MiPaginador(bitacoras, 10)
                    p = 1
                    ppos = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        else:
                            p = paginasesion
                        if 'page_bitacora' in request.GET:
                            p = int(request.GET['page_bitacora'])
                        else:
                            p = paginasesion
                        try:
                            page = pagingbitacora.page(p)
                        except:
                            p = 1
                        page = pagingbitacora.page(p)
                    except:
                        page = pagingbitacora.page(p)
                    # try:
                    #     paginasesionpos = 1
                    #     if 'paginadorpos' in request.session:
                    #         paginasesionpos = int(request.session['paginadorpos'])
                    #     else:
                    #         ppos = paginasesionpos
                    #     if 'page_bitacora_pos' in request.GET:
                    #         ppos = int(request.GET['page_bitacora_pos'])
                    #     else:
                    #         ppos = paginasesionpos
                    #     try:
                    #         pagepos = pagingbitacorapos.page(ppos)
                    #     except:
                    #         ppos = 1
                    #     pagepos = pagingbitacorapos.page(ppos)
                    # except:
                    #     pagepos = paging.page(p)
                    request.session['paginador'] = p
                    data['paging_bitacora'] = pagingbitacora
                    data['searchb'] = searchb if searchb else ""
                    data['aniobit'] = int(aniobit) if aniobit else ""
                    data['mesbit'] = int(mesbit) if mesbit else ""
                    data['rangospaging_bitacora'] = pagingbitacora.rangos_paginado(p)
                    data['page_bitacora'] = page
                    data['bitacoras'] = page.object_list
                    data['bitacora_anios'] = bitacoras.dates('fecha', 'year').distinct()
                    data['reporte_bitacora'] = obtener_reporte('rpt_bitacora')
                    puede_importar_actividades = False
                    # if str(id) in variable_valor('PUEDE_IMPORTAR_ACTIVIDADES_BITACORA'):
                    if puede_realizar_accion_afirmativo(request, 'sga.puede_descargar_db_backup'):
                        puede_importar_actividades = True
                    data['puede_importar_actividades'] = puede_importar_actividades


                    data['administrativo'] = personaadmin = Persona.objects.get(pk=int(request.GET['ida']))
                    data['title'] = u'Hoja de vida'
                    data['datosextension'] = personaadmin.datos_extension()
                    data['perfil'] = personaadmin.mi_perfil()
                    data['examenfisico'] = personaadmin.datos_examen_fisico()
                    data['articulos'] = ArticuloInvestigacion.objects.select_related().filter(participantesarticulos__profesor__persona=personaadmin, status=True,participantesarticulos__status=True).order_by('revista__nombre', 'numero', 'nombre')
                    data['ponencias'] = PonenciasInvestigacion.objects.select_related().filter(participanteponencias__profesor__persona=personaadmin, status=True,participanteponencias__status=True)
                    data['capitulolibro'] = CapituloLibroInvestigacion.objects.select_related().filter(participantecapitulolibros__profesor__persona=personaadmin, status=True,participantecapitulolibros__status=True)
                    data['libros'] = LibroInvestigacion.objects.select_related().filter(participantelibros__profesor__persona=personaadmin, status=True, participantelibros__status=True)
                    data['solicitudes'] = SolicitudPublicacion.objects.filter(persona=personaadmin, aprobado=False,status=True)
                    roles = RolPago.objects.filter(periodo__estado=5, periodo__status=True, persona=personaadmin, status=True)
                    data['reporte_0'] = obtener_reporte('rol_pago')
                    paging = MiPaginador(roles, 25)
                    p = 1
                    try:
                        if 'pagerol' in request.GET:
                            p = int(request.GET['pagerol'])
                        page = paging.page(p)
                    except Exception as ex:
                        page = paging.page(p)
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['roles'] = page.object_list
                    data['niveltitulo'] = NivelTitulacion.objects.filter(status=True).order_by('-rango')
                    activos = ActivoFijo.objects.filter(responsable=personaadmin, statusactivo=1).order_by('ubicacion','descripcion')
                    pagingactivo = MiPaginador(activos, 10)
                    pactivo = 1
                    try:
                        if 'page_activo' in request.GET:
                            pactivo = int(request.GET['page_activo'])
                        pageactivo = pagingactivo.page(pactivo)
                    except Exception as ex:
                        pageactivo = pagingactivo.page(pactivo)
                    data['paging_activo'] = pagingactivo
                    data['rangospaging_activo'] = pagingactivo.rangos_paginado(pactivo)
                    data['page_activo'] = pageactivo
                    data['activos'] = pageactivo.object_list

                    # data['admin'] = personaadmin
                    data['gastospersonales'] = personaadmin.gastospersonales_set.all().order_by('-periodogastospersonales__anio', '-mes')
                    data['pod'] = personaadmin.podevaluaciondet_set.filter(status=True,podperiodo__publicacion__lte=datetime.now().date()).order_by("podperiodo", "departamento")
                    data['certificadostutorias'] = CertificadoTutoriaHV.objects.filter(status=True, persona=personaadmin).order_by('id')
                    # if personaadmin.es_administrativo() or personaadmin.es_profesor():
                    data['articulos'] = ArticuloInvestigacion.objects.select_related().filter((Q(participantesarticulos__profesor__persona=personaadmin) | Q(participantesarticulos__administrativo__persona=personaadmin)), status=True,participantesarticulos__status=True).order_by('-fechapublicacion')
                    data['ponencias'] = PonenciasInvestigacion.objects.select_related().filter((Q(participanteponencias__profesor__persona=personaadmin) | Q(participanteponencias__administrativo__persona=personaadmin)), status=True,participanteponencias__status=True)
                    data['capitulolibro'] = CapituloLibroInvestigacion.objects.select_related().filter((Q(participantecapitulolibros__profesor__persona=personaadmin) | Q(participantecapitulolibros__profesor__persona=personaadmin)), status=True,participantecapitulolibros__status=True)
                    data['libros'] = LibroInvestigacion.objects.select_related().filter((Q(participantelibros__profesor__persona=personaadmin) | Q(participantelibros__profesor__persona=personaadmin)), status=True, participantelibros__status=True)
                    data['solicitudes'] = SolicitudPublicacion.objects.filter(persona=personaadmin, aprobado=False,status=True)
                    distributivos = personaadmin.distributivopersona_set.all()
                    data['anios'] = personaadmin.lista_anios_trabajados_log()
                    data['jornadas'] = personaadmin.historialjornadatrabajador_set.all()
                    if distributivos:
                        data['distributivo'] = distributivo = distributivos[0]
                    else:
                        data['distributivo'] = None
                    # else:
                    #     data['distributivo'] = None
                    #     data['anios'] = None
                    #     data['jornadas'] = None
                    listadocentes = ParticipantesMatrices.objects.values('proyecto__programa__nombre', 'proyecto__nombre', 'proyecto__tipo', 'horas', 'tipoparticipante__nombre').filter( matrizevidencia_id=2, status=True, proyecto__status=True, profesor__persona=personaadmin)
                    listaadministrativo = ParticipantesMatrices.objects.values('proyecto__programa__nombre','proyecto__nombre', 'proyecto__tipo','horas','tipoparticipante__nombre').filter(matrizevidencia_id=2, status=True, proyecto__status=True, administrativo__persona=personaadmin)
                    data['proyectos'] = listadocentes | listaadministrativo
                    data['reporte_1'] = obtener_reporte('hoja_vida_sagest')
                    data['reporte_activos_persona'] = obtener_reporte('activos_persona')
                    data['reporte_capacitaciones_persona'] = obtener_reporte('capacitaciones_persona')
                    data['reporte_certificaciones_persona'] = obtener_reporte('certificaciones_persona')
                    data['reporte_experiencias_persona'] = obtener_reporte('experiencias_persona')
                    data['perfilprincipal'] = perfilprincipal = request.session['perfilprincipal']
                    if personaadmin.profesor_set.filter(status=True).exists():
                        data['es_profesor'] = True
                        profesor = personaadmin.profesor().id
                        data['distributivohoras'] = distributivohoras = personaadmin.profesor().profesordistributivohoras_set.filter(status=True).order_by('-periodo')
                        data['profesor'] = profesor
                        listadodistributivohistorico = DistributivoPersonaHistorial.objects.filter(status=True, persona=personaadmin).distinct().order_by('-pk')
                        listadodistributivoactual = DistributivoPersona.objects.filter(status=True, persona=personaadmin).distinct().order_by('-pk')
                        data['listadodistributivohistorico'] = listadodistributivohistorico
                        data['listadodistributivoactual'] = listadodistributivoactual
                        data['existeactual'] = RespuestaEvaluacionAcreditacion.objects.values('profesor', 'proceso','proceso__mostrarresultados','proceso__periodo','proceso__periodo__nombre').filter(profesor=profesor).distinct().order_by('proceso')
                        data['existe'] = RespuestaEvaluacionAcreditacion.objects.filter(profesor=profesor,tipoinstrumento=1).exists()
                        data['existeanterior'] = ResumenFinalProcesoEvaluacionIntegral.objects.filter(profesor=profesor).exists()
                        data['periodoslect'] = MigracionEvaluacionDocente.objects.values('idperiodo', 'descperiodo','tipoeval').filter(idprofesor=profesor).distinct().order_by('idperiodo')
                        data['periodoslect'] = MigracionEvaluacionDocente.objects.values('idperiodo', 'descperiodo','tipoeval').filter( idprofesor=profesor).distinct().order_by('idperiodo')
                    else:
                        data['es_profesor']=False
                    data['otromeritos'] = OtroMerito.objects.filter(status=True, persona=personaadmin)
                    data['documentopersonal'] = personaadmin.documentos_personales()
                    data['certificaciones'] = CertificacionPersona.objects.filter(status=True, persona=personaadmin)
                    data['reporte_2'] = obtener_reporte('certificado_laboral')
                    data['informesmensuales'] = personaadmin.informemensual_set.filter(status=True).order_by('-fechainicio')
                    data['aportacioneshistoriallaboral'] = personaadmin.personaaportacionhistoriallaboral_set.filter(
                        status=True)

                    return render(request, "th_personal/detallepersonal.html", data)
                except Exception as ex:
                    pass

            elif action == 'descargarcomprimido':
                try:
                    data['form'] = DescargarCompromidoForm()
                    template = get_template("th_hojavida/modal/formcomprimido.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'familiares':
                try:
                    data['title'] = u'Listado de familiares registrados'
                    search = None
                    ids = None
                    search, filtro, url_vars = request.GET.get('s', ''), \
                                                     (Q(status=True)), '&action=familiares'
                    if search:
                        data['s'] = search = request.GET['s'].strip()
                        ss = search.split(' ')
                        url_vars += "&s={}".format(search)
                        if len(ss) == 1:
                            filtro = filtro & (Q(nombres__icontains=search) |
                                               Q(apellido1__icontains=search) |
                                               Q(apellido2__icontains=search) |
                                               Q(cedula__icontains=search) |
                                               Q(pasaporte__icontains=search))
                        else:
                            filtro = filtro & (Q(apellido1__icontains=ss[0]) & Q(apellido2__icontains=ss[1]))
                    elif 'id' in request.GET:
                        ids = request.GET['id']
                        filtro = filtro & (Q(id=ids))

                    distributivo = DistributivoPersona.objects.values_list('persona_id')
                    listado = Persona.objects.filter(filtro,id__in=(distributivo))
                    encuesta = PersonaExtension.objects.filter(persona_id__in=(distributivo)).distinct()
                    data['total'] = total = listado.count()
                    data['total_hijos'] =  encuesta.filter(tienehijos=2).count()
                    data['total_no_hijos'] = encuesta.filter(tienehijos=3).count()
                    data['total_pendiente'] = encuesta.filter(tienehijos=1).count()
                    paging = MiPaginador(listado, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['page'] = page
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['listado'] = page.object_list
                    data['search'] = search if search else ""
                    data['url_vars'] = url_vars
                    data['ids'] = ids if ids else ""
                    return render(request, "th_personal/familiares/familiares.html", data)
                except Exception as ex:
                    pass

            #Información personal

            # Datos personales

            elif action == 'datospersonales':
                try:
                    data['title'] = u'Datos personales'
                    # data['administrativo'] = administrativo = Persona.objects.get(id=int(request.GET['ida']))
                    data['perfil'] = administrativo.mi_perfil()
                    data['migrante'] = MigrantePersona.objects.filter(persona=administrativo).first()
                    data['documentopersonal'] = administrativo.documentos_personales()
                    data['reporte_1'] = obtener_reporte('hoja_vida_sagest')
                    data['reporte_2'] = obtener_reporte('certificado_ficha_estudiantil')
                    request.session['viewactivoth'] = ['informacionpersonal', 'datospersonales']
                    return render(request, "th_personal/informacionpersonal/datospersonales.html", data)
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'editdatospersonales':
                try:
                    data['title'] = u'Datos personales'
                    administrativo = Persona.objects.get(id=encrypt(request.GET['id']))
                    data['id'] = request.GET['id']
                    perfil=administrativo.mi_perfil()
                    form = DatosPersonalesForm(initial={'nombres': administrativo.nombres,
                                                        'apellido1': administrativo.apellido1,
                                                        'apellido2': administrativo.apellido2,
                                                        'cedula': administrativo.cedula,
                                                        'pasaporte': administrativo.pasaporte,
                                                        'extension': administrativo.telefonoextension,
                                                        'sexo': administrativo.sexo,
                                                        'lgtbi': administrativo.lgtbi,
                                                        'anioresidencia': administrativo.anioresidencia,
                                                        'nacimiento': administrativo.nacimiento,
                                                        'email': administrativo.email,
                                                        'estadocivil': administrativo.estado_civil(),
                                                        'libretamilitar': administrativo.libretamilitar,
                                                        'eszurdo': administrativo.eszurdo,
                                                        'estadogestacion': administrativo.estadogestacion,
                                                        'archivocedula':administrativo.documentos_personales().cedula if persona.documentos_personales() else '',
                                                        'papeleta':administrativo.archivo_papeleta() if persona.documentos_personales() else '',
                                                        'archivolibretamilitar':administrativo.archivo_libreta_militar() if persona.documentos_personales() else '',
                                                        'raza': perfil.raza,
                                                        'nacionalidadindigena': perfil.nacionalidadindigena,
                                                        'archivoraza': perfil.archivoraza
                                                        })
                    form.editar()
                    data['form'] = form
                    data['persona'] = persona
                    banderalibreta = 0
                    banderapapeleta = 0
                    banderacedula = 0
                    documentos = PersonaDocumentoPersonal.objects.filter(persona=administrativo)
                    if documentos:
                        if documentos[0].libretamilitar:
                            banderalibreta = 1
                        if documentos[0].papeleta:
                            banderapapeleta = 1
                        if documentos[0].cedula:
                            banderacedula = 1

                    data['banderacedula'] = banderacedula
                    data['banderalibreta'] = banderalibreta
                    data['banderapapeleta'] = banderapapeleta
                    data['switchery']=True
                    template = get_template('th_personal/informacionpersonal/modal/formdatospersonales.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'editdatosnacimiento':
                try:
                    data['title'] = u'Datos de nacimiento'
                    administrativo = Persona.objects.get(id=encrypt(request.GET['id']))
                    data['id'] = request.GET['id']
                    data['nacionalidad'] = administrativo.paisnacimiento.nacionalidad
                    form = DatosNacimientoForm(initial={'paisnacimiento': administrativo.paisnacimiento,
                                                        'provincianacimiento': administrativo.provincianacimiento,
                                                        'cantonnacimiento': administrativo.cantonnacimiento,
                                                        'parroquianacimiento': administrativo.parroquianacimiento,
                                                        'paisnacionalidad': administrativo.paisnacionalidad
                                                        })
                    form.editar(administrativo)
                    data['form'] = form
                    template = get_template('th_personal/informacionpersonal/modal/formnacimiento.html')
                    return JsonResponse({'result':True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'editdatosdomicilio':
                try:
                    data['title'] = u'Datos de domicilio'
                    data['id'] = request.GET['id']
                    administrativo = Persona.objects.get(id=encrypt(request.GET['id']))
                    form = DatosDomicilioForm(initial={'pais': administrativo.pais,
                                                       'provincia': administrativo.provincia,
                                                       'canton': administrativo.canton,
                                                       'ciudadela': administrativo.ciudadela,
                                                       'parroquia': administrativo.parroquia,
                                                       'direccion': administrativo.direccion,
                                                       'direccion2': administrativo.direccion2,
                                                       'sector': administrativo.sector,
                                                       'num_direccion': administrativo.num_direccion,
                                                       'referencia': administrativo.referencia,
                                                       'telefono': administrativo.telefono,
                                                       'telefono_conv': administrativo.telefono_conv,
                                                       'tipocelular': administrativo.tipocelular,
                                                       'archivoplanillaluz':administrativo.archivoplanillaluz,
                                                       'archivocroquis':administrativo.archivocroquis,
                                                       'serviciosbasico':administrativo.documentos_personales().serviciosbasico if persona.documentos_personales() else None,
                                                       'zona': administrativo.zona})
                    form.editar(administrativo)
                    data['form'] = form
                    template = get_template('th_personal/informacionpersonal/modal/formdomicilio.html')
                    return JsonResponse({'result':True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'reportedeclaraciones':
                try:
                    form = DeclaracionesPersonalForm()
                    data['form'] = form
                    data['switchery'] = True
                    template = get_template("th_personal/modal/formreportedeclaraciones.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            # Datos Familiares

            elif action == 'datosfamiliares':
                try:
                    data['title'] = u'Datos familiares'
                    url_vars=f'&action={action}'
                    data['servidor'] = administrativo = Persona.objects.get(id=int(request.GET['ida']))
                    familiares = administrativo.familiares().order_by('-id')
                    paging = MiPaginador(familiares, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['url_vars'] = url_vars
                    data['listado'] = page.object_list
                    request.session['viewactivoth'] = ['informacionpersonal',action]
                    return render(request, "th_personal/informacionpersonal/datosfamiliares.html", data)
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'consultacedula':
                try:
                    cedula = request.GET['cedula'].strip().upper()
                    id=request.GET.get('id',0)
                    datospersona = Persona.objects.filter(Q(pasaporte=cedula) | Q(cedula=cedula) | Q(pasaporte=('VS'+cedula)) | Q(cedula=cedula[2:]), status=True).first()
                    if datospersona:
                        if datospersona == persona:
                            return JsonResponse({"result": "bad", 'mensaje': 'No puede agregar su propia cédula.'})
                        if persona.personadatosfamiliares_set.filter(identificacion=cedula, status=True).exclude(id=id).exists():
                            return JsonResponse({"result": "bad", 'mensaje':'Identificación ingresada ya se encuentra registrada en un familiar'})
                        perfil_i=datospersona.perfilinscripcion_set.filter(status=True).first()
                        editdiscapacidad=False
                        if len(datospersona.mis_perfilesusuarios()) == 1 and datospersona.tiene_usuario_externo():
                            editdiscapacidad = True
                        elif perfil_i and not perfil_i.estadoarchivodiscapacidad == 2:
                            editdiscapacidad = True
                        context={}
                        if perfil_i and perfil_i.tienediscapacidad:
                            context={'tipodiscapacidad':perfil_i.tipodiscapacidad.id if perfil_i.tipodiscapacidad else False,
                                    'tienediscapacidad':perfil_i.tienediscapacidad,
                                    'porcientodiscapacidad':perfil_i.porcientodiscapacidad,
                                    'carnetdiscapacidad':perfil_i.carnetdiscapacidad,
                                    'institucionvalida':perfil_i.institucionvalida.id if perfil_i.institucionvalida else False,
                                    'ceduladiscapacidad':perfil_i.archivo.url if perfil_i.archivo else False ,
                                    'archivoautorizado':perfil_i.archivovaloracion.url if perfil_i.archivovaloracion else False,
                                    }
                        return JsonResponse({"result": "ok",
                                             "apellido1": datospersona.apellido1,
                                             "apellido2": datospersona.apellido2,
                                             "nacimiento": datospersona.nacimiento.strftime('%Y-%m-%d'),
                                             "nombres": datospersona.nombres,
                                             "telefono": datospersona.telefono,
                                             "telefono_conv": datospersona.telefono_conv,
                                             "sexo": datospersona.sexo.id if datospersona.sexo else '' ,
                                             "puedeeditar": editdiscapacidad,
                                             "perfil_i":context})
                    else:
                        return JsonResponse({"result": "no"})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": str(ex)})

            elif action == 'addfamiliar':
                try:
                    form = FamiliarForm()
                    visible_fields = form.visible_fields()
                    total_fields = len(visible_fields)
                    lista=[(1,'Informació Básica', visible_fields[:19]),
                             (2,'Información Laboral', visible_fields[19:27]),
                             (3,'Discapacidad',visible_fields[27:total_fields])
                            ]
                    data['form'] = lista
                    data['id'] = request.GET['id']
                    data['switchery']=True
                    template = get_template('th_personal/informacionpersonal/modal/formfamiliar.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'editfamiliar':
                try:
                    data['title'] = u'Editar familiar'
                    data['id'] = id = request.GET['id']
                    data['idp'] = request.GET['idex']
                    data['familiar'] = familiar = PersonaDatosFamiliares.objects.get(pk=encrypt(id))
                    apellido1, apellido2, nombres, editdiscapacidad, sexo = '', '', familiar.nombre, False, None
                    if familiar.personafamiliar:
                        apellido1 = familiar.personafamiliar.apellido1
                        apellido2 = familiar.personafamiliar.apellido2
                        nombres = familiar.personafamiliar.nombres
                        sexo = familiar.personafamiliar.sexo
                        perfil_f=familiar.personafamiliar.perfilinscripcion_set.filter(status=True).first()
                        estado = True if perfil_f and perfil_f.estadoarchivodiscapacidad == 2 else False
                        if len(familiar.personafamiliar.mis_perfilesusuarios()) == 1 and familiar.personafamiliar.tiene_usuario_externo():
                            editdiscapacidad = True
                        elif not estado:
                            editdiscapacidad = True
                    else:
                        editdiscapacidad = True
                    banderacedula = 0
                    if familiar.cedulaidentidad:
                        banderacedula = 1
                    data['banderacedula'] = banderacedula
                    data['edit_d'] = editdiscapacidad
                    form = FamiliarForm(initial={'identificacion': familiar.identificacion,
                                                 'parentesco': familiar.parentesco,
                                                 'nombre': nombres,
                                                 'apellido1': apellido1,
                                                 'apellido2': apellido2,
                                                 'sexo': sexo,
                                                 'nacimiento': familiar.nacimiento,
                                                 'fallecido': familiar.fallecido,
                                                 'tienediscapacidad': familiar.tienediscapacidad,
                                                 'telefono': familiar.telefono,
                                                 'niveltitulacion': familiar.niveltitulacion,
                                                 'ingresomensual': familiar.ingresomensual,
                                                 'formatrabajo': familiar.formatrabajo,
                                                 'telefono_conv': familiar.telefono_conv,
                                                 'trabajo': familiar.trabajo,
                                                 'convive': familiar.convive,
                                                 'sustentohogar': familiar.sustentohogar,
                                                 'essustituto': familiar.essustituto,
                                                 'autorizadoministerio': familiar.autorizadoministerio,
                                                 'tipodiscapacidad': familiar.tipodiscapacidad,
                                                 'porcientodiscapacidad': familiar.porcientodiscapacidad,
                                                 'carnetdiscapacidad': familiar.carnetdiscapacidad,
                                                 'institucionvalida': familiar.institucionvalida,
                                                 'tipoinstitucionlaboral': familiar.tipoinstitucionlaboral,
                                                 'esservidorpublico': familiar.esservidorpublico,
                                                 'bajocustodia': familiar.bajocustodia,
                                                 'tienenegocio': familiar.tienenegocio,
                                                 'cedulaidentidad': familiar.cedulaidentidad,
                                                 'ceduladiscapacidad': familiar.ceduladiscapacidad,
                                                 'archivoautorizado': familiar.archivoautorizado,
                                                 'cartaconsentimiento': familiar.cartaconsentimiento,
                                                 'archivocustodia': familiar.archivocustodia,
                                                 'centrocuidado': familiar.centrocuidado,
                                                 'centrocuidadodesc': familiar.centrocuidadodesc,
                                                 'negocio': familiar.negocio, })
                    visible_fields = form.visible_fields()
                    total_fields = len(visible_fields)
                    lista = [(1, 'Informació Básica', visible_fields[:19]),
                             (2, 'Información Laboral', visible_fields[19:27]),
                             (3, 'Discapacidad', visible_fields[27:total_fields])
                             ]
                    form.edit()
                    data['form'] = lista
                    data['switchery'] = True
                    template = get_template('th_personal/informacionpersonal/modal/formfamiliar.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            # Datos Financieros

            elif action == 'finanzas':
                try:
                    data['title'] = u'Cuenta bancaria'
                    data['cuentas'] = administrativo.cuentasbancarias()
                    data['declaraciones'] = administrativo.declaracionbienes_set.filter(status=True)
                    request.session['viewactivoth'] = ['informacionpersonal', action]
                    return render(request, "th_personal/informacionpersonal/finanzas.html", data)
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'addcuentabancaria':
                try:
                    form = CuentaBancariaPersonaForm()
                    data['form'] = form
                    data['id'] = request.GET['id']
                    data['switchery'] = True
                    template = get_template('ajaxformmodal.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'editcuentabancaria':
                try:
                    data['title'] = u'Editar cuenta bancaria'
                    data['cuentabancaria'] = cuentabancaria = CuentaBancariaPersona.objects.get(
                        pk=int(request.GET['id']))
                    data['id'] = request.GET['id']
                    data['form'] = CuentaBancariaPersonaForm(initial={'numero': cuentabancaria.numero,
                                                                      'banco': cuentabancaria.banco,
                                                                      'tipocuentabanco': cuentabancaria.tipocuentabanco})
                    template = get_template('ajaxformmodal.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'adddeclaracion':
                try:
                    form = DeclaracionBienForm()
                    data['form'] = form
                    data['id'] = request.GET['id']
                    template = get_template('ajaxformmodal.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'editdeclaracion':
                try:
                    data['title'] = u'Editar declaración de bienes'
                    data['cuentabancaria'] = declaracion = DeclaracionBienes.objects.get(
                        pk=(encrypt(request.GET['id'])))
                    data['id'] = request.GET['id']
                    data['form'] = DeclaracionBienForm(initial=model_to_dict(declaracion))
                    template = get_template('ajaxformmodal.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            # Gastos personales

            elif action == 'gastospersonales':
                try:
                    data['title'] = u'Gastos personales'
                    url_vars = f'&action={action}'
                    gastospersonales = administrativo.gastospersonales_set.filter(status=True).order_by('-periodogastospersonales__anio', '-mes')
                    paging = MiPaginador(gastospersonales, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    formato = PeriodoGastosPersonales.objects.filter(status=True,mostrar=True).first()
                    if formato:
                        data['urlformato'] = formato.formato.url if formato.formato else ''
                    data['url_vars'] = url_vars
                    data['listado'] = page.object_list
                    request.session['viewactivoth'] = ['informacionpersonal', action]
                    return render(request, "th_personal/informacionpersonal/gastospersonales.html", data)
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'subirproyecciongastos':
                try:
                    data['filtro'] = GastosPersonales.objects.get(id=request.GET['id'])
                    template = get_template("th_hojavida/modal/formsubirdoc.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

            elif action == 'detallegasto':
                try:
                    data['gasto'] = gasto = GastosPersonales.objects.get(pk=encrypt_id(request.GET['id']), status=True)
                    data['empleado'] = persona = Persona.objects.get(pk=gasto.persona.id, status=True)
                    data['periodo'] = periodo = PeriodoGastosPersonales.objects.get(pk=gasto.periodogastospersonales.id)
                    data['detalles'] = ResumenMesGastosPersonales.objects.filter(gastospersonales__persona=persona, gastospersonales__periodogastospersonales__anio=periodo.anio).order_by('mes')
                    template = get_template("fin_gastospersonales/detallegasto.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": True, 'data': json_content})
                except Exception as ex:
                    return JsonResponse({"result": False, "mensaje": u"Error al obtener los datos."})

            # FORMULARIO 107

            elif action == 'formulario107':
                try:
                    data['title'] = u'Formulario 107'
                    url_vars = f'&action={action}'
                    formulario107 = administrativo.formulario107_set.filter(status=True).order_by('-anio')
                    paging = MiPaginador(formulario107, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['url_vars'] = url_vars
                    data['listado'] = page.object_list
                    request.session['viewactivoth'] = ['informacionpersonal', action]
                    return render(request, "th_personal/informacionpersonal/formulario107.html", data)
                except Exception as ex:
                    messages.error(request, f'{ex}')


            # Referencias

            elif action == 'referencias':
                try:
                    data['servidor'] = administrativo = Persona.objects.get(id=int(request.GET['ida']))
                    data['title'] = u'Referencias personales'
                    data['listado'] = administrativo.referenciapersona_set.filter(status=True)
                    request.session['viewactivoth'] = ['informacionpersonal', action]
                    return render(request, "th_personal/informacionpersonal/referencias.html", data)
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'addreferencia':
                try:
                    form = ReferenciaPersonaForm()
                    data['form'] = form
                    data['id'] = request.GET['id']
                    data['switchery'] = True
                    template = get_template('ajaxformmodal.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'editreferencia':
                try:
                    data['title'] = u'Editar referencia personal'
                    data['referencia'] = referencia = ReferenciaPersona.objects.get(pk=encrypt(request.GET['id']))
                    form = ReferenciaPersonaForm(initial=model_to_dict(referencia))
                    data['form'] = form
                    data['id'] = request.GET['id']
                    template = get_template('ajaxformmodal.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')


            #Información académica

            # Información titulación

            elif action == 'academica':
                try:
                    data['title'] = u'Formación académica'
                    url_vars = f'&action={action}'
                    data['servidor'] = administrativo = Persona.objects.get(id=int(request.GET['ida']))
                    titulaciones = administrativo.mis_titulaciones().order_by('-titulo__niveltitulacion')
                    paging = MiPaginador(titulaciones, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['url_vars'] = url_vars
                    data['listado'] = page.object_list
                    request.session['viewactivoth'] = ['informacionacademica', action]
                    return render(request, "th_personal/informacionacademica/formacionacademica.html", data)
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'addtitulacion':
                try:
                    data['title'] = u'Adicionar titulación'
                    form = TitulacionPersonaForm()
                    form.adicionar()
                    visible_fields = form.visible_fields()
                    total_fields = len(visible_fields)
                    lista = [(1, 'Información de título', visible_fields[:14]),
                             (2, 'Ubicación de institución', visible_fields[14:18]),
                             (3, 'Beca', visible_fields[18:total_fields])
                             ]
                    data['form'] = lista
                    data['id'] = request.GET['id']
                    data['switchery'] = True
                    template = get_template('th_personal/informacionacademica/modal/formtitulacion.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'edittitulacion':
                try:
                    data['titulacion'] = titulacion = Titulacion.objects.get(pk=encrypt_id(request.GET['id']))

                    form = TitulacionPersonaForm(initial={'titulo': titulacion.titulo,
                                                          'areatitulo': titulacion.areatitulo,
                                                          'fechainicio': titulacion.fechainicio,
                                                          'educacionsuperior': titulacion.educacionsuperior,
                                                          'institucion': titulacion.institucion,
                                                          'colegio': titulacion.colegio,
                                                          'cursando': titulacion.cursando,
                                                          'fechaobtencion': titulacion.fechaobtencion if not titulacion.cursando else datetime.now().date(),
                                                          'fechaegresado': titulacion.fechaegresado if not titulacion.cursando else datetime.now().date(),
                                                          'registro': titulacion.registro,
                                                          # 'areaconocimiento': titulacion.areaconocimiento,
                                                          # 'subareaconocimiento': titulacion.subareaconocimiento,
                                                          # 'subareaespecificaconocimiento': titulacion.subareaespecificaconocimiento,
                                                          'pais': titulacion.pais,
                                                          'provincia': titulacion.provincia,
                                                          'canton': titulacion.canton,
                                                          'parroquia': titulacion.parroquia,
                                                          # 'campoamplio': campoamplio,
                                                          # 'campoespecifico': campoespecifico,
                                                          # 'campodetallado': campodetallado,
                                                          'anios': titulacion.anios,
                                                          'semestres': titulacion.semestres,
                                                          'aplicobeca': titulacion.aplicobeca,
                                                          'tipobeca': titulacion.tipobeca,
                                                          'archivo': titulacion.archivo,
                                                          'registroarchivo': titulacion.registroarchivo,
                                                          'financiamientobeca': titulacion.financiamientobeca,
                                                          'valorbeca': titulacion.valorbeca})
                    form.editar(titulacion)
                    visible_fields = form.visible_fields()
                    total_fields = len(visible_fields)
                    lista = [(1, 'Información de título', visible_fields[:14]),
                             (2, 'Ubicación de institución', visible_fields[14:18]),
                             (3, 'Beca', visible_fields[18:total_fields])
                             ]
                    data['form'] = lista
                    data['id'] = titulacion.id
                    data['idp'] = request.GET['idex']
                    data['switchery'] = True
                    template = get_template('th_hojavida/informacionacademica/modal/formtitulacion.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'buscartitulo':
                try:
                    q = request.GET['q'].upper().strip()
                    s = q.split(" ")
                    filtro = Q(status=True, nombre__icontains=q)
                    titulos = Titulo.objects.filter(filtro).distinct()[:15]
                    resp = [{'id': t.pk, 'text': f"{t.nombre}"} for t in titulos]
                    return JsonResponse(resp, safe=False)
                except Exception as ex:
                    pass

            elif action == 'detalletitulo':
                try:
                    data['titulacion'] = titulacion = Titulacion.objects.get(pk=encrypt_id(request.GET['id']))
                    dettitu = titulacion.detalletitulacionbachiller_set.filter(status=True)
                    data['detalletitulacionbachiller'] = dettitu.last()
                    if titulacion.usuario_creacion:
                        data['personacreacion'] = Persona.objects.get(
                            usuario=titulacion.usuario_creacion) if titulacion.usuario_creacion.id > 1 else ""
                    template = get_template("th_hojavida/detalletitulo.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": True, 'data': json_content})
                except Exception as ex:
                    return JsonResponse({"result": False, "mensaje": u"Error al obtener los datos."})

            elif action == 'adddetalletitulobachiller':
                try:
                    data['titulacion'] = titulacion = Titulacion.objects.get(pk=encrypt_id(request.GET['id']))
                    data['id']=titulacion.id
                    dettitu = titulacion.detalletitulacionbachiller_set.filter(status=True)
                    archivobachiller = Archivo.objects.filter(Q(inscripcion__persona=administrativo)|Q(persona=administrativo),
                                                              tipo_id=16,
                                                              status=True)
                    archivoreconocimiento = Archivo.objects.filter(Q(inscripcion__persona=administrativo)|Q(persona=administrativo),
                                                                   tipo_id=18,
                                                                   status=True)
                    # pasar archivos de detalletitulacion a Archivos y viceversa
                    form = None
                    if dettitu or archivobachiller or archivoreconocimiento:
                        detalle = dettitu.last() if dettitu else None

                        if not detalle and archivobachiller:
                            detalle = DetalleTitulacionBachiller(
                                titulacion=titulacion,
                                actagrado=archivobachiller.last().archivo
                            )
                            detalle.save(request)

                        elif not detalle and archivoreconocimiento:
                            detalle = DetalleTitulacionBachiller(
                                titulacion=titulacion,
                                reconocimientoacademico=archivoreconocimiento.last().archivo
                            )
                            detalle.save(request)

                        elif detalle.actagrado and not archivobachiller:
                            profe = administrativo.perfilprincipal.inscripcion if administrativo.perfilprincipal.es_profesor() else None
                            archivobachiller = Archivo(
                                tipo_id=16,
                                nombre=f'ACTA DE GRADO DE BACHILLER DE LA PERSONA: {administrativo}',
                                fecha=datetime.now().date(),
                                archivo=dettitu.last().actagrado,
                                aprobado=True,
                                profesor=profe,
                                persona=persona,
                                sga=True
                            )
                            archivobachiller.save(request)

                        elif detalle.reconocimientoacademico and not archivoreconocimiento:
                            profe = administrativo.perfilprincipal.inscripcion if administrativo.perfilprincipal.es_profesor() else None
                            archivoreconocimiento = Archivo(
                                tipo_id=18,
                                nombre=f'RECONOCIMIENTO ACADÉMICO DE LA PERSONA: {administrativo}',
                                fecha=datetime.now().date(),
                                archivo=dettitu.last().reconocimientoacademico,
                                aprobado=True,
                                profesor=profe,
                                persona=persona,
                                sga=True
                            )
                            archivoreconocimiento.save(request)

                        form = DetalleTitulacionBachillerForm(
                            initial={'titulacion': titulacion.id if titulacion else "",
                                     'calificacion': detalle.calificacion,
                                     'archivo': detalle.actagrado,
                                     'anioinicioperiodograduacion': detalle.anioinicioperiodograduacion,
                                     'aniofinperiodograduacion': detalle.aniofinperiodograduacion,
                                     'reconocimientoacademico': detalle.reconocimientoacademico})
                    else:
                        form = DetalleTitulacionBachillerForm(
                            initial={'titulacion': titulacion.id if titulacion else ""})
                    # form.deshabilitar(titulacion)

                    data['form'] = form
                    template = get_template('th_personal/informacionacademica/modal/formtitulacionbachiller.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            #Articulos

            elif action == 'articulos':
                try:
                    data['title'] = u'Artículos'
                    url_vars = f'&action={action}'

                    # Consulta los artículos
                    articulos = ArticuloInvestigacion.objects.select_related().filter((Q(
                        participantesarticulos__profesor__persona=administrativo) | Q(
                        participantesarticulos__administrativo__persona=administrativo) | Q(
                        participantesarticulos__inscripcion__persona=administrativo)),
                        status=True, aprobado=True,
                        participantesarticulos__status=True).order_by('-fechapublicacion')

                    paging = MiPaginador(articulos, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['url_vars'] = url_vars
                    data['listado'] = page.object_list
                    request.session['viewactivoth'] = ['informacionacademica', action]
                    return render(request, "th_personal/informacionacademica/articulos.html", data)
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'evidenciasarticulo':
                try:
                    data['articulo'] = articulos = ArticuloInvestigacion.objects.get(pk=int(encrypt(request.GET['id'])))
                    data['evidencias'] = Evidencia.objects.filter(status=True, matrizevidencia_id=3)

                    template = get_template("th_hojavida/detallearticulo.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    msg = ex.__str__()
                    return JsonResponse({"result": "bad", "message": u"Error al obtener los datos. [%s]" % msg})

            #Capacitaciones


            elif action == 'capacitaciones':
                try:
                    data['title'] = u'Capacitaciones'
                    url_vars, tipo = f'&action={action}&ida={administrativo.pk}',request.GET.get('tipo','')
                    url_vars += f'&tipo={tipo}'
                    if tipo == 'eventos':
                        listado = administrativo.distributivopersona_set.filter(status=True, estadopuesto=PUESTO_ACTIVO_ID)
                    elif tipo == 'solicitudes':
                        listado = administrativo.lista_evento_realizado_persona().order_by('-fechasolicitud')
                    elif tipo == 'perfeccionamiento':
                        listado = CapCabeceraSolicitudDocente.objects.filter(participante=administrativo, notificado=True, rutapdf__isnull=False).order_by('-fechasolicitud')
                    else:
                        listado = administrativo.mis_capacitaciones().order_by('-fechafin')
                    paging = MiPaginador(listado, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['url_vars'] = url_vars
                    data['tipo']=tipo
                    data['solicitud_capacitacion'] = variable_valor('SOLICITUD_CAPACITACION')
                    data['pendiente_capacitacion'] = variable_valor('PENDIENTE_CAPACITACION')
                    data['aprobado_capacitacion'] = variable_valor('APROBADO_CAPACITACION')
                    data['reporte_capacitaciones_persona'] = obtener_reporte('capacitaciones_persona')
                    data['listado'] = page.object_list
                    request.session['viewactivoth'] = ['informacionacademica', action]
                    return render(request, "th_personal/informacionacademica/capacitaciones.html", data)
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'detallecapacitacion':
                try:
                    data['capacitacion'] = capacitacion = Capacitacion.objects.get(pk=int(request.GET['id']))
                    if capacitacion.usuario_creacion:
                        data['personacreacion'] = Persona.objects.get(
                            usuario=capacitacion.usuario_creacion) if capacitacion.usuario_creacion.id > 1 else ""
                    template = get_template("th_hojavida/detallecapacitacion.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, "mensaje": f"Error al obtener los datos.{ex}"})

            elif action == 'addcapacitacion':
                try:
                    data['title'] = u'Adicionar capacitación'
                    form = CapacitacionPersonaForm()
                    form.adicionar()
                    visible_fields = form.visible_fields()
                    total_fields = len(visible_fields)
                    lista = [(1, 'Datos del evento', visible_fields[:8]),
                             (2, 'Datos de la capacitación', visible_fields[8:21]),
                             (3, 'Datos de ubicación', visible_fields[21:total_fields])
                             ]
                    data['form'] = lista
                    data['id'] = request.GET['id']
                    template = get_template('th_personal/informacionacademica/modal/formcapacitaciones.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'editcapacitacion':
                try:
                    data['title'] = u'Editar capacitación'
                    id=encrypt_id(request.GET['id'])
                    data['capacitacion'] = capacitacion = Capacitacion.objects.get(pk=id)
                    data['modalidad1'] = capacitacion.modalidad
                    form = CapacitacionPersonaForm(initial={'institucion': capacitacion.institucion,
                                                            'nombre': capacitacion.nombre,
                                                            'descripcion': capacitacion.descripcion,
                                                            'tipo': capacitacion.tipo,
                                                            'tipocurso': capacitacion.tipocurso,
                                                            'tipocertificacion': capacitacion.tipocertificacion,
                                                            'tipocapacitacion': capacitacion.tipocapacitacion,
                                                            'tipoparticipacion': capacitacion.tipoparticipacion,
                                                            'auspiciante': capacitacion.auspiciante,
                                                            'expositor': capacitacion.expositor,
                                                            'anio': capacitacion.anio,
                                                            'contexto': capacitacion.contextocapacitacion,
                                                            'detallecontexto': capacitacion.detallecontextocapacitacion,
                                                            'areaconocimiento': capacitacion.areaconocimiento,
                                                            'subareaconocimiento': capacitacion.subareaconocimiento,
                                                            'subareaespecificaconocimiento': capacitacion.subareaespecificaconocimiento,
                                                            'pais': capacitacion.pais,
                                                            'provincia': capacitacion.provincia,
                                                            'canton': capacitacion.canton,
                                                            'parroquia': capacitacion.parroquia,
                                                            'fechainicio': capacitacion.fechainicio,
                                                            'fechafin': capacitacion.fechafin,
                                                            'horas': capacitacion.horas,
                                                            'modalidad': capacitacion.modalidad,
                                                            'archivo': capacitacion.archivo,
                                                            'otramodalidad': capacitacion.otramodalidad}, instancia=capacitacion)
                    form.editar(capacitacion)
                    visible_fields = form.visible_fields()
                    total_fields = len(visible_fields)
                    lista = [(1, 'Datos del evento', visible_fields[:8]),
                             (2, 'Datos de la capacitación', visible_fields[8:21]),
                             (3, 'Datos de ubicación', visible_fields[21:total_fields])]
                    data['form'] = lista
                    data['id']=id
                    template = get_template('th_personal/informacionacademica/modal/formcapacitaciones.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'detalleeventoaprobacion_docente':
                try:
                    data['cabecerasolicitud'] = cabecera = CapCabeceraSolicitudDocente.objects.get(
                        pk=int(request.GET['id']))
                    data['detallesolicitud'] = cabecera.capdetallesolicituddocente_set.all()
                    data['solicitud_capacitacion'] = variable_valor('SOLICITUD_CAPACITACION')
                    data['pendiente_capacitacion'] = variable_valor('PENDIENTE_CAPACITACION')
                    data['aprobado_capacitacion'] = variable_valor('APROBADO_CAPACITACION')
                    template = get_template("th_hojavida/detalleeventoaprobacion.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": True, 'data': json_content})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

            elif action == 'detalleeventoaprobacion':
                try:
                    data['cabecerasolicitud'] = cabecera = CapCabeceraSolicitud.objects.get(pk=int(request.GET['id']))
                    data['detallesolicitud'] = cabecera.capdetallesolicitud_set.all()
                    data['solicitud_capacitacion'] = variable_valor('SOLICITUD_CAPACITACION')
                    data['pendiente_capacitacion'] = variable_valor('PENDIENTE_CAPACITACION')
                    data['aprobado_capacitacion'] = variable_valor('APROBADO_CAPACITACION')
                    template = get_template("th_hojavida/detalleeventoaprobacion.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": True, 'data': json_content})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": False, "mensaje": u"Error al obtener los datos."})

            elif action == 'detalleevento':
                try:
                    data['evento'] = evento = CapEventoPeriodo.objects.get(pk=int(request.GET['id']))
                    template = get_template("th_hojavida/detalleevento.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": True, 'data': json_content})
                except Exception as ex:
                    return JsonResponse({"result": False, "mensaje": u"Error al obtener los datos."})


            # CERTIFICACIONES

            elif action == 'certificaciones':
                try:
                    data['title'] = u'Certificaciones internacionales'
                    url_vars, tipo = f'&action={action}&ida={administrativo.id}', request.GET.get('tipo','certificacionidiomas')
                    if tipo == 'certificacioninternacionales':
                        certificados = CertificacionPersona.objects.filter(status=True, persona=administrativo).order_by('-id')
                        data['reporte_certificaciones_persona'] = obtener_reporte('certificaciones_persona')
                    elif tipo == 'certificacionidiomas':
                        certificados = CertificadoIdioma.objects.filter(status=True, persona=administrativo).order_by('-id')
                        data['reporte_certificaciones_internacional'] = obtener_reporte('certificaciones_idiomas')
                    url_vars+=f'&tipo={tipo}'
                    paging = MiPaginador(certificados, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['url_vars'] = url_vars
                    data['t_certipersona']=len(CertificacionPersona.objects.filter(status=True, persona=administrativo).values_list('id'))
                    data['t_certiidioma']=len(CertificadoIdioma.objects.filter(status=True, persona=administrativo).values_list('id'))
                    data['tipo'] = tipo
                    data['listado'] = page.object_list
                    request.session['viewactivoth'] = ['informacionacademica', action]
                    return render(request, "th_personal/informacionacademica/certificaciones.html", data)
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'addcertificadoidioma':
                try:
                    form = CertificadoIdiomaForm()
                    data['switchery']=True
                    data['form'] = form
                    data['id'] = request.GET['id']
                    template = get_template('th_personal/informacionacademica/modal/formcertificadoidioma.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'editcertificadoidioma':
                try:
                    certificado = CertificadoIdioma.objects.get(pk=encrypt_id(request.GET['id']))
                    data['form'] = CertificadoIdiomaForm(initial=model_to_dict(certificado))
                    data['switchery'] = True
                    data['id']=certificado.id
                    template = get_template('th_personal/informacionacademica/modal/formcertificadoidioma.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'detalleotroscertificacion':
                try:
                    data['certificacion'] = certificacion = CertificadoIdioma.objects.get(pk=encrypt(request.GET['id']))
                    if certificacion.usuario_creacion:
                        data['personacreacion'] = Persona.objects.get(
                            usuario=certificacion.usuario_creacion) if certificacion.usuario_creacion.id > 1 else ""
                    template = get_template("th_hojavida/detalleotroscertificacion.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, "mensaje": f"Error: {ex}"})

            elif action == 'addcertificado':
                try:
                    data['title'] = u'Adicionar Certificación'
                    form = CertificadoPersonaForm()
                    data['form'] = form
                    data['id'] = request.GET['id']
                    data['switchery'] = True
                    template = get_template('th_personal/informacionacademica/modal/formcertificado.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'editcertificado':
                try:
                    data['title'] = u'Editar Certificación'
                    data['certificado'] = certificado = CertificacionPersona.objects.get(pk=encrypt_id(request.GET['id']))
                    initial = model_to_dict(certificado)
                    form = CertificadoPersonaForm(initial=initial)
                    form.editar(certificado)
                    data['form'] = form
                    data['switchery'] = True
                    data['id'] = certificado.id
                    template = get_template('th_personal/informacionacademica/modal/formcertificado.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'detallecertificacion':
                try:
                    data['certificacion'] = certificacion = CertificacionPersona.objects.get(pk=int(request.GET['id']))
                    if certificacion.usuario_creacion:
                        data['personacreacion'] = Persona.objects.get(
                            usuario=certificacion.usuario_creacion) if certificacion.usuario_creacion.id > 1 else ""
                    template = get_template("th_hojavida/detallecertificacion.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, "mensaje": f"Error: {ex}"})


            # Discapacidad

            elif action == 'discapacidad':
                try:
                    data['title'] = u'Discapacidad'
                    data['subtitle'] = u'Información cargada'
                    data['url_vars'] = f'&action={action}'
                    data['perfil'] = administrativo.mi_perfil()
                    data['documentopersonal'] = administrativo.documentos_personales()
                    request.session['viewactivoth'] = ['informacionmedica',action]
                    return render(request, "th_personal/informacionmedica/discapacidad.html", data)
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'editdiscapacidad':
                try:
                    from sga.models import Discapacidad
                    data['title'] = u'Discapacidad'
                    data['id'] = id = request.GET['id']
                    administrativo= Persona.objects.get(id=encrypt(id))
                    perfil = persona.mi_perfil()
                    form = DiscapacidadForm(initial=model_to_dict(perfil))
                    tienearchivo = True if perfil.archivo else False
                    data['form'] = form
                    data['tienearchivo'] = tienearchivo
                    data['switchery'] = True
                    data['tipodis'] = Discapacidad.objects.filter(status=True)
                    template = get_template('th_personal/informacionmedica/modal/formdiscapacidad.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            # Datos médicos
            elif action == 'datosmedicos':
                try:
                    data['title'] = u'Datos médicos'
                    data['subtitle'] = u'Información cargada'
                    data['url_vars'] = f'&action={action}'
                    data['perfil'] = administrativo.mi_perfil()
                    data['datosextension'] =datosextension = administrativo.datos_extension()
                    data['examenfisico'] =examenfisico = administrativo.datos_examen_fisico()
                    data['vacunascovid'] = VacunaCovid.objects.filter(persona=administrativo, status=True).order_by('-id')
                    request.session['viewactivoth'] = ['informacionmedica',action]
                    return render(request, "th_personal/informacionmedica/datosmedicos.html", data)
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'editdatosmedicos':
                try:
                    administrativo = Persona.objects.get(id=encrypt(request.GET['id']))
                    datosextension = administrativo.datos_extension()
                    examenfisico = administrativo.datos_examen_fisico()
                    form = DatosMedicosForm(initial={
                        'carnetiess' : datosextension.carnetiess,
                        'sangre' : administrativo.sangre,
                        'archivotiposangre': administrativo.archivo_tiposangre ,
                        'peso' : examenfisico.peso,
                        'talla' : examenfisico.talla
                    })
                    data['form'] = form
                    data['id'] = request.GET['id']
                    template = get_template('th_personal/informacionmedica/modal/formbasico.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'editcontactoemergencia':
                try:
                    administrativo = Persona.objects.get(id=encrypt(request.GET['id']))
                    datosextension = administrativo.datos_extension()
                    form = ContactoEmergenciaForm(initial={
                        'contactoemergencia' : datosextension.contactoemergencia,
                        'parentescoemergencia' : datosextension.parentescoemergencia,
                        'telefonoemergencia' : datosextension.telefonoemergencia,
                        'telefonoconvemergencia': datosextension.telefonoconvemergencia,
                        'correoemergencia': datosextension.correoemergencia
                    })
                    data['id'] = request.GET['id']
                    data['form'] = form
                    template = get_template('th_hojavida/informacionmedica/modal/formbasico.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')



            #Información Laboral

            # Contratos

            elif action == 'contratos':
                try:
                    data['title'] = u'Contratos'
                    url_vars=f'&action={action}'
                    data['servidor'] = administrativo = Persona.objects.get(id=int(request.GET['ida']))
                    contratos = administrativo.mis_contratos()
                    paging = MiPaginador(contratos, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['url_vars'] = url_vars
                    data['listado'] = page.object_list
                    request.session['viewactivoth'] = ['informacionlaboral',action]
                    return render(request, "th_personal/informacionlaboral/contratos.html", data)
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'addcontrato':
                try:
                    data['title'] = u'Adicionar contrato'
                    data['form'] = PersonaContratosForm()
                    data['id'] = request.GET['id']
                    template = get_template('ajaxformmodal.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'descargarcontratos':
                try:
                    data['form'] = DescargarContratosForm()
                    data['id'] = request.GET['id']
                    template = get_template('ajaxformmodal.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'editcontrato':

                try:

                    data['title'] = u'Editar contrato'

                    filtro = PersonaContratos.objects.get(pk=int(encrypt(request.GET['id'])))

                    data['form'] = PersonaContratosForm(initial=model_to_dict(filtro))

                    data['id'] = request.GET['id']

                    template = get_template('ajaxformmodal.html')

                    return JsonResponse({"result": True, 'data': template.render(data)})

                except Exception as ex:

                    messages.error(request, f'{ex}')

            elif action == 'detallecontrato':
                try:
                    data['title'] = u'Detalles contrato'
                    data['filtro'] = PersonaContratos.objects.get(pk=int(encrypt(request.GET['id'])))
                    template = get_template('th_personal/informacionlaboral/modal/infocontrato.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            # Acción de personal

            elif action == 'accionpersonal':
                try:
                    data['title'] = u'Acción de personal'
                    url_vars=f'&action={action}&ida={administrativo.id}'
                    data['servidor'] = administrativo = Persona.objects.get(id=int(request.GET['ida']))
                    acciones = administrativo.accionpersonal_set.filter(status=True).order_by('-fechadesde').exclude(motivoaccion_id=6)
                    acciones2 = administrativo.personaacciones_set.filter(status=True, migrado=False).order_by('-fecharige')
                    paging = MiPaginador(acciones, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['url_vars'] = url_vars
                    data['listado'] = page.object_list
                    data['listado2'] = acciones2
                    request.session['viewactivoth'] = ['informacionlaboral',action]
                    return render(request, "th_personal/informacionlaboral/accionpersonal.html", data)
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'accionpersonalvacaciones':
                try:
                    data['title'] = u'Acción de personal'
                    url_vars=f'&action={action}&ida={administrativo.id}'
                    data['servidor'] = administrativo = Persona.objects.get(id=int(request.GET['ida']))
                    acciones = administrativo.accionpersonal_set.filter(status=True, motivoaccion_id=6).order_by('-id')
                    # acciones2 = administrativo.personaacciones_set.filter(status=True, migrado=False, motivo_id=6).order_by('-id')
                    paging = MiPaginador(acciones, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['url_vars'] = url_vars
                    data['listado'] = page.object_list
                    # data['listado2'] = acciones2
                    request.session['viewactivoth'] = ['informacionlaboral','accionpersonal']
                    return render(request, "th_personal/informacionlaboral/accionpersonal.html", data)
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'addaccion':
                try:
                    data['title'] = u'Adicionar acción de personal'
                    data['form'] = PersonaAccionesForm()
                    data['id'] = request.GET['id']
                    template = get_template('ajaxformmodal.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'editaccionpersonal':
                try:
                    data['title'] = u'Editar acción de personal'
                    accion = AccionPersonal.objects.get(pk=int(encrypt(request.GET['id'])))
                    data['form'] = PersonaAccionesForm(initial={
                        'numerodocumento': accion.numerodocumento,
                        'tipo':accion.tipo,
                        'motivoaccion':accion.motivoaccion,
                        'denominacionpuesto':accion.denominacionpuesto,
                        'unidadorganica':accion.departamento,
                        'remuneracion':accion.rmu,
                        'explicacion':accion.explicacion,
                        'fechainicio':accion.fechadesde,
                        'documento':accion.documento,
                        'archivo':accion.archivo
                    })
                    data['id'] = request.GET['id']
                    template = get_template('ajaxformmodal.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'editaccionpersonal_old':
                try:
                    data['title'] = u'Editar acción de personal'
                    filtro = PersonaAcciones.objects.get(pk=int(encrypt(request.GET['id'])))
                    data['form'] = PersonaAccionesOldForm(initial=model_to_dict(filtro))
                    data['id'] = request.GET['id']
                    template = get_template('ajaxformmodal.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'detallaccionpersonal':
                try:
                    data['title'] = u'Detalles de acción de personal'
                    data['filtro'] = AccionPersonal.objects.get(pk=int(encrypt(request.GET['id'])))
                    template = get_template('th_personal/informacionlaboral/modal/infoaccionpersonal.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'detallaccionpersonal_old':
                try:
                    data['title'] = u'Detalles de acción de personal'
                    data['filtro'] = PersonaAcciones.objects.get(pk=int(encrypt(request.GET['id'])))
                    template = get_template('th_personal/informacionlaboral/modal/infoaccionpersonal.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            # Situación laboral

            elif action == 'situacionlaboral':
                try:
                    data['title'] = u'Situación laboral'
                    url_vars = f'&action={action}&ida={administrativo.id}'
                    historiallaboral = administrativo.personaaportacionhistoriallaboral_set.filter(status=True)
                    paging = MiPaginador(historiallaboral, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['url_vars'] = url_vars
                    data['listado'] = page.object_list
                    request.session['viewactivoth'] = ['informacionlaboral', action]
                    return render(request, "th_personal/informacionlaboral/situacionlaboral.html", data)
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'addaportacionhistoriallaboral':
                try:
                    data['form'] = PersonaAportacionHistorialLaboralForm()
                    data['idp'] = valor = request.GET['idex']
                    data['id'] = id = request.GET['id']
                    if valor =='edit':
                        data['registro'] = registro = PersonaAportacionHistorialLaboral.objects.filter(pk=encrypt(id)).first()
                        data['form'] = PersonaAportacionHistorialLaboralForm(initial=model_to_dict(registro))
                    template = get_template('ajaxformmodal.html')
                    json_content = template.render(data, request=request)
                    return JsonResponse({"result": True, 'data': json_content})
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al cargar los datos."})


            # Rol de pago

            elif action == 'rolpago':
                try:
                    data['title'] = u'Roles de pago'
                    url_vars=f'&action={action}&ida={administrativo.id}'
                    roles = administrativo.rolpago_set.filter(periodo__estado=5, periodo__status=True, status=True)
                    paging = MiPaginador(roles, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['url_vars'] = url_vars
                    data['listado'] = page.object_list
                    data['reporte_0'] = obtener_reporte('rol_pago')

                    request.session['viewactivoth'] = ['informacionlaboral',action]
                    return render(request, "th_personal/informacionlaboral/rolpago.html", data)
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'detallerol':
                    try:
                        data['detallerol'] = registro = RolPago.objects.get(pk=int(request.GET['id']), status=True)
                        data['detalleinformativo'] = registro.detallerolinformativo()
                        data['detalleingreso'] = registro.detallerolingreso()
                        data['detalleegreso'] = registro.detallerolegreso_consolidado()
                        template = get_template("th_nomina/detallerol.html")
                        json_content = template.render(data)
                        return JsonResponse({"result": True, 'data': template.render(data)})
                    except Exception as ex:
                        messages.error(request, f'{ex}')

            # Datos institución

            elif action == 'datosinstitucion':
                try:
                    data['title'] = u'Datos institucionales'
                    data['subtitle'] = u'Información cargada'
                    data['url_vars'] = f'&action={action}&ida={administrativo.id}'
                    data['perfil'] = administrativo.mi_perfil()
                    data['documentopersonal'] = administrativo.documentos_personales()
                    request.session['viewactivoth'] = ['informacionlaboral',action]
                    return render(request, "th_personal/informacionlaboral/datosinstitucion.html", data)
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'editdatosinstitucion':
                try:
                    data['title'] = u'Editar datos de institución'
                    filtro = personaadmin = Persona.objects.get(pk=int(encrypt(request.GET['id'])))
                    data['form'] = DatosPersonaInstitucionForm(initial={'indicebiometrico': personaadmin.identificacioninstitucion,
                                                                        'registro': personaadmin.regitrocertificacion,
                                                                        'servidorcarrera': personaadmin.servidorcarrera,
                                                                        'extension': personaadmin.telefonoextension,
                                                                        'correoinstitucional': personaadmin.emailinst,
                                                                        'fechaingresoies': personaadmin.fechaingresoies if personaadmin.fechaingresoies else datetime.now().date(),
                                                                        'fechasalidaies': personaadmin.fechasalidaies if personaadmin.fechasalidaies else datetime.now().date(),
                                                                        'labora': True if not personaadmin.fechasalidaies else False,
                                                                        'concursomeritos': personaadmin.concursomeritos})
                    data['id'] = request.GET['id']
                    template = get_template('ajaxformmodal.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')


            # Evaluación

            elif action == 'evaluacion':
                try:
                    data['title'] = u'Evaluaciones'
                    url_vars, tipo = f'&action={action}&ida={administrativo.id}', request.GET.get('tipo', '')
                    data['profesor'] = profesor = administrativo.profesor()

                    if not tipo:
                        listado = administrativo.podevaluaciondet_set.filter(status=True, podperiodo__publicacion__lte=datetime.now().date()).order_by( "podperiodo", "departamento")
                    elif tipo == 'docentepregrado':
                        url_vars += f'&tipo={tipo}'
                        data['existeanterior'] = ResumenFinalProcesoEvaluacionIntegral.objects.filter(profesor=profesor).exists()
                        data['periodoslect'] = MigracionEvaluacionDocente.objects.values('idperiodo', 'descperiodo', 'tipoeval').filter(idprofesor=profesor.id).distinct().order_by('idperiodo')
                        listado = RespuestaEvaluacionAcreditacion.objects.filter(proceso__mostrarresultados=True).values('profesor', 'profesor__pk', 'proceso', 'proceso__pk',
                                                                                'proceso__mostrarresultados', 'proceso__periodo',
                                                                                'proceso__periodo__id',
                                                                                'proceso__periodo__nombre').filter(profesor=profesor.id).distinct().order_by('proceso')
                    elif tipo == 'docenteposgrado':
                        url_vars += f'&tipo={tipo}'
                        listado = RespuestaEvaluacionAcreditacion.objects.filter(proceso__mostrarresultados=False).values('profesor', 'proceso', 'proceso__pk',
                                                                     'proceso__mostrarresultados', 'proceso__periodo',
                                                                     'proceso__periodo__nombre').filter(profesor=profesor.id).distinct().order_by('proceso')
                    paging = MiPaginador(listado, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['tipo'] = tipo
                    data['url_vars'] = url_vars
                    data['listado'] = page.object_list
                    request.session['viewactivoth'] = ['informacionlaboral',action]
                    return render(request, "th_personal/informacionlaboral/evaluacion.html", data)
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'pdfmodeloactual':
                try:
                    data['title'] = u'Adicionar Programa'
                    adduserdata(request, data)
                    data['departamento'] = departamento = Departamento.objects.get(pk=128)
                    data['profesor'] = profesor = administrativo.profesor().id
                    data['administrativo'] = administrativo
                    data['datospersona'] = administrativo
                    data['profesor'] = profesor
                    data['procesoperiodo'] = ProcesoEvaluativoAcreditacion.objects.get(periodo=request.GET['idperiodo'], status=True)
                    if ResponsableEvaluacion.objects.filter(status=True, activo=True).exists():
                        data['responsable'] = ResponsableEvaluacion.objects.filter(status=True, activo=True)[0]
                    data['nomperiodo'] = request.GET['nomperiodo']
                    data['resultados'] = porcentaje = ResumenFinalEvaluacionAcreditacion.objects.get(distributivo__profesor=profesor,distributivo__periodo=request.GET['idperiodo'])
                    data['porcentaje'] = notaporcentaje = round(((porcentaje.resultado_total * 100) / 5), 2)
                    data['fechactual'] = datetime.now().strftime("%d") + '/' + datetime.now().strftime("%m") + '/' + datetime.now().strftime("%y")+ ' ' + datetime.now().strftime("%H:%M")
                    notaporcen = str(request.GET['idperiodo']) + "-" + administrativo.cedula + "-" + str(notaporcentaje)
                    qrname = 'qrce_mied_' + request.GET['idperiodo'] + administrativo.cedula
                    # folder = SITE_STORAGE + 'media/qrcode/evaluaciondocente/'
                    folder = os.path.join(os.path.join(SITE_STORAGE, 'media', 'qrcode', 'evaluaciondocente', 'qr'))
                    # folder = os.path.join(SITE_STORAGE, 'media', 'qrcode', 'evaluaciondocente')
                    rutapdf = folder + qrname + '.pdf'
                    rutaimg = folder + qrname + '.png'
                    if os.path.isfile(rutapdf):
                        os.remove(rutaimg)
                        os.remove(rutapdf)
                    url = pyqrcode.create('http://sga.unemi.edu.ec//media/qrcode/evaluaciondocente/' + qrname + '.pdf')
                    imageqr = url.png(folder + qrname + '.png', 16, '#000000')
                    imagebarcode = code128.image(notaporcen).save(folder + qrname + "_bar.png")
                    data['qrname'] = 'qr' + qrname
                    data['url_path'] = 'http://127.0.0.1:8000'
                    if not DEBUG:
                        data['url_path'] = 'https://sga.unemi.edu.ec'
                    return conviert_html_to_pdfsave(
                        'pro_certificados/pdfqrce_modeloactual.html',
                        {
                            'pagesize': 'A4',
                            'listadoevaluacion': data,
                        },qrname + '.pdf'
                    )
                except Exception as ex:
                    pass

            # Marcadas

            elif action == 'marcadas':
                try:
                    data['title'] = u'LOG de Marcadas'
                    # data['distributivo'] = distributivo = DistributivoPersona.objects.get(pk=int(encrypt(request.GET['id'])))
                    data['anios'] = administrativo.lista_anios_trabajados_log()
                    data['jornadas'] = administrativo.historialjornadatrabajador_set.all()
                    puede_crear_marcada = False
                    try:
                        puede_crear_marcada = puede_realizar_accion_is_superuser(request, 'sagest.puede_crear_marcada')
                    except:
                        puede_crear_marcada = False
                    data['puede_crear_marcada'] = puede_crear_marcada
                    data['destino'] = 'th_hojavida'
                    data['pued_modificar'] = 1
                    # if not persona.id == persona.id:
                    #     raise NameError('Error')
                    data['hora'] = str(datetime.now().time())[0:5]
                    request.session['viewactivoth'] = ['informacionlaboral', action]
                    return render(request, "th_personal/informacionlaboral/marcadas.html", data)
                except Exception as ex:
                    pass

            # Plantilla

            elif action == 'plantillahis':
                try:
                    data['title'] = u'Plantilla'
                    url_vars=f'&action={action}&ida={administrativo.id}'
                    plantilla = administrativo.distributivopersonahistorial_set.filter(status=True).distinct().order_by('-id')
                    paging = MiPaginador(plantilla, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['url_vars'] = url_vars
                    data['listado'] = page.object_list
                    data['reporte_0'] = obtener_reporte('rol_pago')

                    request.session['viewactivoth'] = ['informacionlaboral',action]
                    return render(request, "th_personal/informacionlaboral/plantilla.html", data)
                except Exception as ex:
                    messages.error(request, f'{ex}')

            # Experiencia

            elif action == 'experiencia':
                try:
                    data['title'] = u'Experiencian laboral'
                    url_vars, filtro, search, = f'&action={action}&ida={administrativo.id}', \
                                                          Q(status=True, persona=administrativo), \
                                                          request.GET.get('s', '').strip()
                    if search:
                        filtro = filtro & Q(Q(titulo__icontains=search))
                        data['s'] = search
                        url_vars += f'&s={search}'
                    experiencias = ExperienciaLaboral.objects.filter(filtro)
                    paging = MiPaginador(experiencias, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['url_vars'] = url_vars
                    data['listado'] = page.object_list
                    data['reporte_experiencias_persona'] = obtener_reporte('experiencias_persona')
                    request.session['viewactivoth'] = ['informacionlaboral', action]
                    return render(request, "th_personal/informacionlaboral/experiencialaboral.html", data)
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'addexperiencia':
                try:
                    data['title'] = u'Adicionar experiencia'
                    data['id'] = request.GET['id']
                    form = ExperienciaLaboralForm()
                    form.adicionar()
                    visible_fields = form.visible_fields()
                    total_fields = len(visible_fields)
                    lista = [(1, 'Información básica', visible_fields[:9]),
                             (2, 'Ubicación empresarial', visible_fields[9:13]),
                             (3, 'Datos laborales', visible_fields[13:total_fields])
                             ]
                    data['form'] = lista
                    data['switchery'] = True
                    template = get_template('th_personal/informacionlaboral/modal/formexperiencia.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'editexperiencia':
                try:
                    data['experiencia'] = experiencia = ExperienciaLaboral.objects.get(pk=encrypt_id(request.GET['id']))
                    data['id']=experiencia.id
                    vigente = True
                    fechafin = ""
                    if experiencia.motivosalida:
                        vigente = False
                        fechafin = experiencia.fechafin
                    form = ExperienciaLaboralForm(initial={'tipoinstitucion': experiencia.tipoinstitucion,
                                                                   'institucion': experiencia.institucion,
                                                                   'cargo': experiencia.cargo,
                                                                   'departamento': experiencia.departamento,
                                                                   'pais': experiencia.pais,
                                                                   'provincia': experiencia.provincia,
                                                                   'canton': experiencia.canton,
                                                                   'parroquia': experiencia.parroquia,
                                                                   'fechainicio': experiencia.fechainicio,
                                                                   'fechafin': fechafin,
                                                                   'motivosalida': experiencia.motivosalida,
                                                                   'vigente': vigente,
                                                                   'regimenlaboral': experiencia.regimenlaboral,
                                                                   'horassemanales': experiencia.horassemanales,
                                                                   'dedicacionlaboral': experiencia.dedicacionlaboral,
                                                                   'actividadlaboral': experiencia.actividadlaboral,
                                                                   'observaciones': experiencia.observaciones})
                    visible_fields = form.visible_fields()
                    total_fields = len(visible_fields)
                    lista = [(1, 'Información básica', visible_fields[:9]),
                             (2, 'Ubicación empresarial', visible_fields[9:13]),
                             (3, 'Datos laborales', visible_fields[13:total_fields])
                             ]
                    data['form'] = lista
                    data['switchery'] = True
                    template = get_template('th_personal/informacionlaboral/modal/formexperiencia.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            # Tutorias Posgrado

            elif action == 'sustentaciones':
                try:
                    data['title'] = u"Seguimiento de tutorías posgrado"
                    temas_tribunal = TribunalTemaTitulacionPosgradoMatricula.objects.filter(Q(status=True),
                                                                                            Q(presidentepropuesta__persona=administrativo) |
                                                                                            Q(secretariopropuesta__persona=administrativo) |
                                                                                            Q(delegadopropuesta__persona=administrativo)).order_by(
                        '-fechadefensa', '-horadefensa')
                    sustentaciones = []
                    for tribunal in temas_tribunal:
                        if tribunal.tematitulacionposgradomatriculacabecera:
                            participante = tribunal.tematitulacionposgradomatriculacabecera.obtener_parejas()[0]
                        else:
                            participante = tribunal.tematitulacionposgradomatricula

                        detallecalificacion = participante.calificaciontitulacionposgrado_set.filter(
                            status=True).order_by('tipojuradocalificador')
                        if detallecalificacion.exists():
                            promediopuntajetrabajointegral = \
                                detallecalificacion.values_list('puntajetrabajointegral').aggregate(
                                    promedio=Avg('puntajetrabajointegral'))['promedio']
                            promediodefensaoral = detallecalificacion.values_list('puntajedefensaoral').aggregate(
                                promedio=Avg('puntajedefensaoral'))['promedio']
                            promediofinal = detallecalificacion.values_list('puntajerubricas').aggregate(
                                promedio=Avg('puntajerubricas'))['promedio']
                            if str(participante.mecanismotitulacionposgrado_id) in variable_valor(
                                    'ID_MECANISMO_ARTICULOS'):
                                if promediofinal >= 70:
                                    sustentaciones.append(tribunal)
                            else:
                                if promediofinal > 0:
                                    sustentaciones.append(tribunal)
                    data['sustentaciones'] = sustentaciones
                    request.session['viewactivoth'] = ['seguimiento', 'sustentaciones']
                    return render(request, "th_personal/seguimiento/viewsustentacion.html", data)
                except Exception as ex:
                    line_err = f"Error en la linea {sys.exc_info()[-1].tb_lineno}"
                    err_ = f"Ocurrio un error, {ex.__str__()}. {line_err}"
                    return HttpResponseRedirect(f"{request.path}?info={err_}")

            return HttpResponseRedirect(request.path)
        else:
            try:
                data['title'] = u'Personal de la institución'
                search = None
                ids = None
                tipopersonal = None
                tipopersonal, search, filtro, url_vars = request.GET.get('tipopersonal', '1'), \
                                                         request.GET.get('s', ''), \
                                                         (Q(status=True) & Q(administrativo__isnull=False) |
                                                          Q(profesor__isnull=False)), ''
                if search:
                    data['s'] = search = request.GET['s'].strip()
                    ss = search.split(' ')
                    url_vars += "&s={}".format(search)
                    if len(ss) == 1:
                        filtro = filtro & (Q(nombres__icontains=search) |
                                           Q(apellido1__icontains=search) |
                                           Q(apellido2__icontains=search) |
                                           Q(cedula__icontains=search) |
                                           Q(pasaporte__icontains=search))
                    else:
                        filtro = filtro &(Q(apellido1__icontains=ss[0]) & Q(apellido2__icontains=ss[1]))
                elif 'id' in request.GET:
                    ids = request.GET['id']
                    filtro = filtro & (Q(id=ids))

                tipopersonal = 1 if 'tipopersonal' not in request.GET else int(request.GET['tipopersonal'])

                if tipopersonal==1:
                    data['tipopersonal'] = int(tipopersonal)
                    filtro = filtro & (Q(distributivopersona__estadopuesto_id=1, distributivopersona__status=True))
                url_vars += "&tipopersonal={}".format(tipopersonal)
                administrativos = Persona.objects.filter(filtro).distinct().order_by('apellido1')

                data['total'] = administrativos.count()
                data['es_director_th'] = es_director_th = persona.es_directordepartamental_talentohumano()
                paging = MiPaginador(administrativos, 25)
                p = 1
                try:
                    paginasesion = 1
                    if 'paginador' in request.session:
                        paginasesion = int(request.session['paginador'])
                    if 'page' in request.GET:
                        p = int(request.GET['page'])
                    else:
                        p = paginasesion
                    try:
                        page = paging.page(p)
                    except:
                        p = 1
                    page = paging.page(p)
                except:
                    page = paging.page(p)
                request.session['paginador'] = p
                data['paging'] = paging
                data['rangospaging'] = paging.rangos_paginado(p)
                data['page'] = page
                # data['search'] = search if search else ""
                # data['ids'] = ids if ids else ""
                data['administrativos'] = page.object_list
                data['grupo_docentes'] = PROFESORES_GROUP_ID
                data['grupo_empleadores'] = EMPLEADORES_GRUPO_ID
                data['grupo_estudiantes'] = ALUMNOS_GROUP_ID
                data['grupo_administrativos'] = variable_valor('ADMINISTRATIVOS_GROUP_ID')
                data['titulos_nuevos'] = Persona.objects.filter(Q(administrativo__isnull=False) | Q(profesor__isnull=False), titulacion__verificado=False, titulacion__cursando=False).distinct().count()
                data['cursos_nuevos'] = Persona.objects.filter(Q(administrativo__isnull=False) | Q(profesor__isnull=False), capacitacion__verificado=False).distinct().count()
                data['declaraciones_nuevos'] = Persona.objects.filter(Q(administrativo__isnull=False) | Q(profesor__isnull=False), declaracionbienes__verificado=False).distinct().count()
                data['cbancarias_nuevos'] = Persona.objects.filter(Q(administrativo__isnull=False) | Q(profesor__isnull=False), cuentabancariapersona__verificado=False).distinct().count()
                data['experiencia_nuevos'] = Persona.objects.filter(Q(administrativo__isnull=False) | Q(profesor__isnull=False), experiencialaboral__verificado=False).distinct().count()
                data['total_nuevos'] = 0
                data['periodosgastos'] = PeriodoGastosPersonales.objects.filter(status=True)
                data['tipopersonal'] = tipopersonal
                data['url_vars'] = url_vars
                return render(request, 'th_personal/view.html', data)
            except Exception as ex:
                pass

def validar_choque_horario_actividad_gestion(persona, fi, hi, hf, periodo=None):
    listaturnosocupados, choqueturno = [], []
    dia = fi.isoweekday()
    if Profesor.objects.filter(persona=persona, status=True, activo=True).values('id').exists():
        # HORARIO DE DOCENCIA
        profesor = Profesor.objects.filter(persona=persona, status=True, activo=True).order_by('-id').first()
        fActiv = Q(status=True, activo=True, inicio__lte=fi, fin__gte=fi, turno__status=True, dia=dia, detalledistributivo__distributivo__profesor=profesor)
        fClass = Q(materia__visiblehorario=True, materia__fechafinasistencias__gte=fi, status=True, activo=True, inicio__lte=fi, fin__gte=fi, turno__mostrar=True, turno__status=True, dia=dia, materia__profesormateria__profesor_id=profesor.id)
        if periodo:
            if not periodo.es_posgrado() and not periodo.es_admision():
                fClass &= Q(materia__nivel__periodo=periodo)
                fActiv &= Q(detalledistributivo__distributivo__periodo=periodo)
        profesormateria = ProfesorMateria.objects.values_list('materia_id', flat=True).filter(materia__inicio__lte=fi, materia__fin__gte=fi, profesor=profesor, profesor__status=True, activo=True,principal=True, status=True, materia__status=True,materia__cerrado=False).exclude(materia__nivel__periodo__tipo__id__in=[3, 4]).distinct()
        for m in Materia.objects.filter(pk__in=profesormateria):
            clases = m.clase_set.filter(fClass).order_by('turno__comienza')
            for clase in clases:
                listaturnosocupados.append([clase.turno, m])
                if clases.filter(Q(id=clase.id) & (Q(turno__comienza__range=(hi, hf)) | Q(turno__termina__range=(hi, hf)))):
                    choqueturno.append(clase.turno.id)
        # HORARIO DE ACTIVIDADES
        actividades = ClaseActividad.objects.filter(fActiv).exclude(detalledistributivo__criteriodocenciaperiodo__criterio__procesotutoriaacademica=True).distinct().order_by('turno__comienza')
        ids_choque = actividades.values_list('id', flat=True).filter(Q(turno__comienza__range=(hi, hf)) | Q(turno__termina__range=(hi, hf)))
        for ac in actividades:
            actividadnombre = ''
            if ac.detalledistributivo.criteriodocenciaperiodo:
                actividadnombre = ac.detalledistributivo.criteriodocenciaperiodo.criterio.nombre
            if ac.detalledistributivo.criterioinvestigacionperiodo:
                actividadnombre = ac.detalledistributivo.criterioinvestigacionperiodo.criterio.nombre
            if ac.detalledistributivo.criteriogestionperiodo:
                actividadnombre = ac.detalledistributivo.criteriogestionperiodo.criterio.nombre
            if ac.id in ids_choque:
                choqueturno.append(ac.turno.id)

            listaturnosocupados.append([ac.turno, actividadnombre])

    return [choqueturno, listaturnosocupados]


def descargar_comprimido_requisitos(request, ePersonPeriodo):
    # Crear directorios recursivamente
    directory_p = os.path.join(MEDIA_ROOT, 'talento_humano')
    directory = os.path.join(directory_p, 'docs_comprimidos')
    os.makedirs(directory_p, exist_ok=True)
    os.makedirs(directory, exist_ok=True)
    name_folder = f'{unidecode(ePersonPeriodo.persona.nombre_completo_minus())} (Requisitos de ingreso)'
    name_zip = f'{name_folder}.zip'
    url_zip = os.path.join(SITE_STORAGE, 'media', 'talento_humano', 'docs_comprimidos', name_zip)
    fantasy_zip = zipfile.ZipFile(url_zip, 'w')
    for idx, r in enumerate(ePersonPeriodo.documentos_subidos()):
        url_file = r.archivo.url
        name_file = unidecode(r.requisito.nombre.replace(" ", "_"))
        ext = url_file[url_file.rfind("."):].lower()
        ruta_archivo_zip = os.path.join(name_folder, f'{name_file}{ext}')
        fantasy_zip.write(r.archivo.path, ruta_archivo_zip)
    fantasy_zip.close()
    response = HttpResponse(open(url_zip, 'rb'), content_type='application/zip')
    response['Content-Disposition'] = f'attachment; filename={name_zip}'
    return response