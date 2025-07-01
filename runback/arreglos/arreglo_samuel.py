#!/usr/bin/env python
# -*- coding: utf-8 -*-
import os
import sys
import openpyxl
from django.contrib.auth.hashers import make_password
from django.db.models import Q

# SITE_ROOT = os.path.dirname(os.path.realpath(__file__))
YOUR_PATH = os.path.dirname(os.path.realpath(__file__))
# print(f"YOUR_PATH: {YOUR_PATH}")
SITE_ROOT = os.path.dirname(os.path.dirname(YOUR_PATH))
SITE_ROOT = os.path.join(SITE_ROOT, '')
# print(f"SITE_ROOT: {SITE_ROOT}")
your_djangoproject_home = os.path.split(SITE_ROOT)[0]
# print(f"your_djangoproject_home: {your_djangoproject_home}")
sys.path.append(your_djangoproject_home)

from django.core.wsgi import get_wsgi_application

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "settings")
application = get_wsgi_application()

from sga.tasks import send_html_mail, conectar_cuenta
from datetime import datetime
from sga.models import Persona, CUENTAS_CORREOS, DetalleDistributivo, CriterioDocenciaPeriodo, Periodo, CriterioDocencia
from pdip.models import ActividadesPerfil, ContratoDip, PerfilPuestoDip, ActividadesContratoPerfil, \
    Departamento as DepContrato, Gestion
from settings import SITE_STORAGE
from postulate.models import Partida, PersonaAplicarPartida
from sga.funciones import null_to_decimal


def migraractividades():
    contador = 0
    actividades = ActividadesPerfil.objects.filter(status=True)
    contratos = ContratoDip.objects.filter(status=True)
    cargos = PerfilPuestoDip.objects.filter(status=True)
    for cont in contratos:
        cargos_actividades = cont.cargo.actividades.all()
        contratos_actividades = cont.actividadesextra.all()

        # Uno las actividades del contrato y perfil
        actividades_unidas = cargos_actividades | contratos_actividades
        actividades_unidas = actividades_unidas.filter(status=True).distinct()

        for acti in actividades_unidas:
            if not ActividadesContratoPerfil.objects.filter(status=True, contrato=cont, perfil=cont.cargo,
                                                            actividad=acti).exists():
                actcontperf = ActividadesContratoPerfil(
                    contrato=cont,
                    perfil=cont.cargo,
                    actividad=acti
                )
                actcontperf.save()
                contador += 1

    #     for act in cont.actividadesextra.all():
    #
    #         if not ActividadesContratoPerfil.objects.filter(status=True,contrato = cont,actividad=act).exists():
    #             activi = ActividadesContratoPerfil(
    #                 contrato = cont,
    #                 actividad= act,
    #             )
    #             activi.save()
    #             contador+=1
    #             print("Contrato: {} actividades: {}".format(cont.persona,act))
    # for carg in cargos:
    #     for act in carg.actividades.all():
    #         if not ActividadesContratoPerfil.objects.filter(status=True, perfil=carg, actividad=act).exists():
    #             activi = ActividadesContratoPerfil(
    #                 perfil = carg,
    #                 actividad= act,
    #             )
    #             activi.save()
    #             contador += 1
    #             print("Cargo: {} actividades: {}".format(carg.nombre, act))
    print(f'Total de registros afectados: {contador}')


# migraractividades()

from helpdesk.models import HdGruposCategoria, HdMantenimientoGruCategoria, HdMantenimientoGruDanios, \
    HdTareasActivosPreventivos, HdTareasActivosPreventivosDanios, HdPiezaPartes as hd

from sagest.models import GruposCategoria, MantenimientoGruCategoria, MantenimientoGruDanios, \
    TareasActivosPreventivosDanios, TareasActivosPreventivos, HdPiezaPartes as sg, \
    ActivoTecnologico, Rubro, Departamento as DepPdip, InscritoCongreso
from django.db import transaction, connections
from ejecuform.adm_ejecuform import buscarPedidoOnlineRubroEpunemi
from ejecuform.models import CapaEventoInscritoFormaEjecutiva
from sagest.adm_capacitacioneventoperiodoipec import buscarPagosEpunemiRubroUnemi


def miggrupoactividades():
    try:
        # grupocategoria = GruposCategoria.objects.filter(status=True).order_by('id')
        # cont = 0
        # for cat in grupocategoria:
        #     if not HdGruposCategoria.objects.values('id').filter(id = cat.pk).exists():
        #         hd_grupocat = HdGruposCategoria(
        #             id = cat.pk,
        #             status= cat.status,
        #             descripcion=cat.descripcion
        #         )
        #         cont +=1
        #         hd_grupocat.save()
        #         print(F"{cont}.- {hd_grupocat.descripcion} - cod: {hd_grupocat.pk}")
        # print('Se migro los datos de la tabla sagest.GruposCategoria a la tabla helpdesk.HdGruposCategoria')
        # print(F'Un total de {cont} registros')
        #
        # cont = 0
        #
        # mantgrucat = MantenimientoGruCategoria.objects.filter(status=True)
        #
        # for mantgru in mantgrucat:
        #     if not HdMantenimientoGruCategoria.objects.values('id').filter(id=mantgru.pk).exists():
        #         hd_mantgrucat = HdMantenimientoGruCategoria(
        #             id = mantgru.pk,
        #             status= mantgru.status,
        #             fecha_creacion= mantgru.fecha_creacion,
        #             fecha_modificacion= mantgru.fecha_modificacion,
        #             descripcion= mantgru.descripcion,
        #             grupocategoria=HdGruposCategoria.objects.get(pk=mantgru.grupocategoria.pk),
        #             usuario_creacion= mantgru.usuario_creacion,
        #             usuario_modificacion=mantgru.usuario_modificacion,
        #             activo=mantgru.activo
        #         )
        #         cont += 1
        #         hd_mantgrucat.save()
        #         print(F"{cont}.- {hd_mantgrucat.descripcion} - cod: {hd_mantgrucat.pk}")
        # print('Se migro los datos de la tabla sagest.MantenimientoGruCategoria a la tabla helpdesk.HdMantenimientoGruCategoria')
        # print(F'Un total de {cont} registros')
        #
        # cont = 0
        #
        # mantgrudan = MantenimientoGruDanios.objects.filter(status = True)
        #
        # for mantdan in mantgrudan:
        #     if not HdMantenimientoGruDanios.objects.values('id').filter(id=mantdan.pk).exists():
        #         hd_mantgrudani = HdMantenimientoGruDanios(
        #             id = mantdan.pk,
        #             status= mantdan.status,
        #             fecha_creacion=mantdan.fecha_creacion,
        #             descripcion=mantdan.descripcion,
        #             activo = mantdan.activo,
        #             grupocategoria=HdGruposCategoria.objects.get(pk=mantdan.grupocategoria.pk),
        #             usuario_creacion=mantdan.usuario_creacion,
        #         )
        #         cont += 1
        #         hd_mantgrudani.save()
        #         print(F"{cont}.- {hd_mantgrudani.descripcion} - cod: {hd_mantgrudani.pk}")
        # print('Se migro los datos de la tabla sagest.MantenimientoGruCategoria a la tabla helpdesk.HdMantenimientoGruCategoria')
        # print(F'Un total de {cont} registros')
        #
        # cont = 0
        # piezasagest = sg.objects.filter(status=True)
        # for pieza in piezasagest:
        #     if not hd.objects.filter(id = pieza.pk).exists():
        #         hd_piezas = hd(
        #             id = pieza.pk,
        #             descripcion = pieza.descripcion,
        #             status = pieza.status,
        #             imagen = pieza.imagen,
        #             grupocategoria= HdGruposCategoria.objects.get(id=pieza.grupocategoria.pk) if pieza.grupocategoria else None
        #         )
        #         cont += 1
        #         hd_piezas.save()
        #         print(F"{cont}.- {hd_piezas.descripcion} - cod: {hd_piezas.pk}")
        # print('Se migro los datos de la tabla sagest.HdPiezaPartes a la tabla helpdesk.HdPiezaPartes')
        # print(F'Un total de {cont} registros')

        activotec = ActivoTecnologico.objects.filter(status=True, activotecnologico__catalogo__grupo__isnull=False)
        cont = 0
        for act in activotec:
            if not act.tipoactivo:
                tipo = act.activotecnologico.catalogo.grupo
                act.tipoactivo = HdGruposCategoria.objects.get(id=tipo.pk)
                cont += 1
                act.save()
                print(F"{cont}.- {act.__str__()} - cod: {act.pk}")
        print('Se Actualizo la tabla sagest.ActivoTecnologico, migrando los datos de la categoria del activo')
        print(F'Un total de {cont} registros')
    except Exception as ex:
        transaction.set_rollback(True)
        print(F'Ha ocurrido el siguiente error: {ex.__str__()}')


# miggrupoactividades()


from sga.models import PracticasPreprofesionalesInscripcion, Profesor


def cambiartutorsupervisor():
    ruta = os.path.join(SITE_ROOT, 'runback', 'arreglos', 'media')
    # excel = openpyxl.load_workbook("{}media/cambio_tutores_16102022.xlsx".format(SITE_STORAGE))
    # excel = openpyxl.load_workbook("{}/media/cambio_tutores_20102022.xlsx".format(SITE_STORAGE))
    excel = openpyxl.load_workbook("{}{}cambio_tutores_11012023.xlsx".format(ruta, os.sep))
    # lista = excel['Hoja1']
    lista = excel['Hoja4']
    numlist = lista.rows
    cab = 0
    cont = 0
    cont2 = 0
    listacod = []
    for pract in numlist:
        cab += 1
        supervisor = None
        tutor = None
        if cab > 1:
            if Profesor.objects.filter(persona__cedula=pract[10].value, status=True).exists():
                supervisor = Profesor.objects.filter(persona__cedula=pract[10].value, status=True).order_by(
                    '-id').first()
            if Profesor.objects.filter(persona__cedula=pract[8].value, status=True).exists():
                tutor = Profesor.objects.filter(persona__cedula=pract[8].value, status=True).order_by('-id').first()
            if PracticasPreprofesionalesInscripcion.objects.filter(pk=int(pract[0].value), status=True).exists():
                if supervisor and tutor:
                    practicas = PracticasPreprofesionalesInscripcion.objects.filter(pk=int(pract[0].value),
                                                                                    status=True).order_by('-id').first()
                    print(f"Tutor actual: {practicas.tutorunemi} - Tutor nuevo: {tutor}")
                    print(f"Supervisor actual: {practicas.supervisor} - Supervisor nuevo: {supervisor}")
                    practicas.tutorunemi = tutor
                    practicas.supervisor = supervisor
                    practicas.save()
                    cont += 1
                    print(f"No. {cont} - {practicas.__str__()} - tutor: {tutor} - supervisor: {supervisor}")
                else:
                    listacod.append(pract[0].value)
                    cont2 += 1
    print(f"Total registro actualizados: {cont}")
    print(f"Total registro sin actualizar: {cont2}")
    print(f"Cods de registro sin actualizar: {listacod}")


# cambiartutorsupervisor()

from sga.models import Insignia, InsigniaPersona


def agregarinsignia():
    ruta = os.path.join(SITE_ROOT, 'runback', 'arreglos', 'media')
    excel = openpyxl.load_workbook("{}{}desfile_17092022.xlsx".format(ruta, os.sep))

    insignia = Insignia.objects.get(id=3, status=True)
    listaNoexiste = []
    listaInsigniaExiste = []
    cont = 0
    cont2 = 0
    cont3 = 0
    for i in excel.sheetnames[2:]:
        lista = excel[i]
        numlist = lista.rows
        cab = 0
        for personaIn in numlist:
            cab += 1
            personaI = None
            if cab > 1:
                if Persona.objects.filter(cedula=personaIn[20].value).order_by('-id').exists():
                    personaInsign = Persona.objects.filter(cedula=personaIn[20].value).order_by('-id').first()
                    if not InsigniaPersona.objects.filter(persona=personaInsign, insignia=insignia,
                                                          status=True).order_by('-id').exists():
                        insiniapersonanueva = InsigniaPersona(
                            persona=personaInsign,
                            insignia=insignia,
                            fechaobtencion=datetime(2022, 9, 17).date()
                        )
                        insiniapersonanueva.save()
                        cont += 1
                        print(f'No. {cont} {insiniapersonanueva} - {personaInsign} - {insignia}')
                    else:
                        cont2 += 1
                        listaInsigniaExiste.append({'cedula': personaIn[20].value, 'nombre': personaIn[21].value})
                else:
                    cont3 += 1
                    listaNoexiste.append({'cedula': personaIn[20].value, 'nombre': personaIn[21].value})
    print(f'Total de registro ingresados: {cont}')
    print(f'Total de personas que no se encontraron ({cont3}) {listaNoexiste}')
    print(f'Total de personas que ya tienen la insignia({cont2}) {listaInsigniaExiste}')


# agregarinsignia()


from sga.models import PreInscripcionPracticasPP, Inscripcion


def eliminarinscripcionppp():
    try:
        cont = 0
        from sga.templatetags.sga_extras import encrypt
        from django.db.models import Q
        lista = []
        idperiodo = encrypt('OPPQQRRSSTTUUVVWWXWS')
        preinscripcion = PreInscripcionPracticasPP.objects.get(status=True, pk=int(idperiodo))
        inscripciones = Inscripcion.objects.filter(detallepreinscripcionpracticaspp__preinscripcion=preinscripcion,
                                                   status=True).distinct()
        for inscripcion in inscripciones:
            listapre = inscripcion.detallepreinscripcionpracticaspp_set.values_list('itinerariomalla_id',
                                                                                    flat=False).filter(status=True,
                                                                                                       estado=1,
                                                                                                       preinscripcion=preinscripcion)
            if inscripcion.inscripcionmalla_set.values('id').exists():
                if inscripcion.inscripcionmalla_set.filter(status=True)[0].malla.itinerariosmalla_set.filter(
                        status=True).exists():
                    matricula = inscripcion.matricula_set.filter(status=True)[0]
                    listaitinerariorealizado = inscripcion.cumple_total_horas_itinerario()
                    itinerariosvalidosid = []
                    for it in inscripcion.inscripcionmalla_set.filter(status=True)[0].malla.itinerariosmalla_set.filter(
                            status=True):
                        nivelhasta = it.nivel.orden
                        if inscripcion.todas_materias_aprobadas_rango_nivel2(1, nivelhasta):
                            itinerariosvalidosid.append(it.pk)
                    itinerarios = inscripcion.inscripcionmalla_set.filter(status=True)[
                        0].malla.itinerariosmalla_set.values_list('id', flat=True).filter(status=True,
                                                                                          nivel__orden__lte=matricula.nivelmalla.orden).filter(
                        pk__in=itinerariosvalidosid).exclude(id__in=listaitinerariorealizado)
                    itinerariosculminados = PracticasPreprofesionalesInscripcion.objects.values_list(
                        'preinscripcion__itinerariomalla__id', flat=True).filter(status=True, inscripcion=inscripcion,
                                                                                 preinscripcion__itinerariomalla__id__in=itinerarios,
                                                                                 culminada=True)
                    itinerarios = itinerariosculminados.union(itinerarios)
                    deleteinscripcion = inscripcion.detallepreinscripcionpracticaspp_set.filter(status=True, estado=1,
                                                                                                preinscripcion=preinscripcion).exclude(
                        itinerariomalla_id__in=itinerarios)  # .exclude(itinerariomalla_id__in=listapre) #~Q(itinerariomalla_id__in = itinerarios),
                    for pre in deleteinscripcion:
                        cont += 1
                        lista.append(pre.id)
                        pre.status = False
                        pre.save()
                        print(
                            f'{cont}. Se borro la preinscripcion ({pre.id}) - {pre} {pre.itinerariomalla} del estudiante {pre.inscripcion}')

        print(f'Total de registro de detallepreinscripcion de practicas borrados: {cont}')
        print(f'Lista de id de los registros afectados: {lista}')
    except Exception as ex:
        print(ex.__str__())


# eliminarinscripcionppp()

# CASO MUÑOZ LEAL
# def eliminarpractica():
#     try:
#         cont = 0
#         persona = Persona.objects.filter(status=True, cedula='0940930423').order_by('-id').first()
#         from sga.templatetags.sga_extras import encrypt
#         from django.db.models import Q
#         lista = []
#         idperiodo = encrypt('OPPQQRRSSTTUUVVWWXWS')
#         preinscripcion = PreInscripcionPracticasPP.objects.get(status=True, pk=int(idperiodo))
#         inscripcion = Inscripcion.objects.filter(detallepreinscripcionpracticaspp__preinscripcion=preinscripcion, status=True,persona=persona)[0]
#         listapre = inscripcion.detallepreinscripcionpracticaspp_set.values_list('itinerariomalla_id', flat=False).filter(status=True, estado=1, preinscripcion=preinscripcion)
#         if inscripcion.inscripcionmalla_set.values('id').exists():
#             if inscripcion.inscripcionmalla_set.filter(status=True)[0].malla.itinerariosmalla_set.filter(status=True).exists():
#                 matricula = inscripcion.matricula_set.filter(status=True)[0]
#                 listaitinerariorealizado = inscripcion.cumple_total_horas_itinerario()
#                 itinerariosvalidosid = []
#                 for it in inscripcion.inscripcionmalla_set.filter(status=True)[0].malla.itinerariosmalla_set.filter(status=True):
#                     nivelhasta = it.nivel.orden
#                     if inscripcion.todas_materias_aprobadas_rango_nivel2(1, nivelhasta):
#                         itinerariosvalidosid.append(it.pk)
#                 itinerarios = inscripcion.inscripcionmalla_set.filter(status=True)[0].malla.itinerariosmalla_set.values_list('id', flat=True).filter(status=True, nivel__orden__lte=matricula.nivelmalla.orden).filter(pk__in=itinerariosvalidosid).exclude(id__in=listaitinerariorealizado)
#                 itinerariosculminados = PracticasPreprofesionalesInscripcion.objects.values_list('preinscripcion__itinerariomalla__id', flat=True).filter(status=True, inscripcion=inscripcion, preinscripcion__itinerariomalla__id__in=itinerarios, culminada=True)
#                 itinerarios = itinerariosculminados.union(itinerarios)
#                 deleteinscripcion = inscripcion.detallepreinscripcionpracticaspp_set.filter(status=True, estado=1, preinscripcion=preinscripcion).exclude(itinerariomalla_id__in=itinerarios)  # .exclude(itinerariomalla_id__in=listapre) #~Q(itinerariomalla_id__in = itinerarios),
#                 for pre in deleteinscripcion:
#                     cont += 1
#                     lista.append(pre.id)
#                     pre.status = False
#                     pre.save()
#                     print(f'{cont}. Se borro la preinscripcion ({pre.id}) - {pre} {pre.itinerariomalla} del estudiante {pre.inscripcion}')
#
#         print(f'Total de registro de detallepreinscripcion de practicas borrados: {cont}')
#         print(f'Lista de id de los registros afectados: {lista}')
#     except Exception as ex:
#         print(ex.__str__())
# eliminarpractica()

def eliminar_rubros_formacionejecutica():
    with transaction.atomic():
        try:
            capevento_id = CapaEventoInscritoFormaEjecutiva.objects.filter(status=True).values_list('id', flat=True)
            rubros = Rubro.objects.filter(status=True, cancelado=False,
                                          capeventoperiodoformejecu__isnull=False,
                                          capeventoperiodoformejecu_id__in=capevento_id)
            rubrospagados_noeliminados, total_rubrosunemi, total_rubrosepunemi = [], 0, 0
            contador = 0
            for r in rubros:
                contador += 1
                # consultar rubro epunemi
                cursor = connections['epunemi'].cursor()
                sql = """SELECT id, valor, totalunemi, cancelado, nombre, persona_id FROM sagest_rubro WHERE idrubrounemi=%s AND status=TRUE; """ % (
                    r.id)
                cursor.execute(sql)
                rubroepunemi = cursor.fetchone()
                inscrito = CapaEventoInscritoFormaEjecutiva.objects.filter(status=True,
                                                                           capeventoperiodo_id=r.capeventoperiodoformejecu.id,
                                                                           participante=r.persona).first()
                if rubroepunemi:
                    tienepagosepunemi = buscarPagosEpunemiRubroUnemi(rubroepunemi[0])
                    tienepagosepunemi_min = Rubro.objects.values_list('idrubroepunemi', flat=True).filter(
                        capeventoperiodoformejecu=r.capeventoperiodoformejecu, persona=r.persona, status=True)
                    for tpago in tienepagosepunemi_min:
                        tienepagosepunemi = buscarPagosEpunemiRubroUnemi(tpago)
                        if tienepagosepunemi:
                            break
                    if not r.tiene_pagos() and r.cancelado == False and not tienepagosepunemi and rubroepunemi[
                        3] == False and r.valor == rubroepunemi[1] and r.valor == rubroepunemi[2]:
                        # Verificar que no tenga pendiente registro de pago, es decir que no tenga comprobante alumno epunemi
                        rubroepunemitienecomprobante = buscarPedidoOnlineRubroEpunemi(rubroepunemi[5], rubroepunemi[0])
                        if not rubroepunemitienecomprobante:
                            # Elimino el rubro en UNEMI y llamo funcion eliminar en EPUNEMI
                            r.status = False
                            r.save()
                            total_rubrosunemi += 1
                            sql = """SELECT row_to_json(r) FROM (SELECT * FROM sagest_rubro WHERE status=TRUE AND id=%s) r;"""
                            cursor.execute(sql, [rubroepunemi[0]])
                            rubroepunemi_dic = cursor.fetchone()
                            rubroepunemi_dic = rubroepunemi_dic[0]

                            # Elimino logicamente el rubro en Epunemi
                            sql = """UPDATE sagest_rubro SET status=false WHERE status=TRUE AND id=%s; """ % (
                                rubroepunemi[0])
                            cursor.execute(sql)
                            total_rubrosepunemi += 1

                            print(
                                f"{contador}.- Se elimino rubro de unemi: [{r.persona.cedula} - {r.id}], y rubro de epunemi: [{rubroepunemi[0]}].")
                        else:
                            rubrospagados_noeliminados.append(
                                f'{r.persona.cedula} [{r.id}] porque el rubro posee un comprobante de pago')
                            print(
                                f"{contador}.- No se pudo eliminar el rubro: {r.persona.cedula} [{r.id}] porque el rubro posee un comprobante de pago")
                    else:
                        rubrospagados_noeliminados.append(
                            f'{r.persona.cedula} [{r.id}] porque el rubro posee pagos o su valor difiere con el rubro EPUNEMI')
                        print(
                            f"{contador}.- No se pudo eliminar el rubro: {r.persona.cedula} [{r.id}] porque el rubro posee pagos o su valor difiere con el rubro EPUNEM")
                else:
                    # Elimino el rubro que sólo existe en UNEMI
                    # guardar auditoría en UNEMI el log del rubro elimminado en unemi
                    r.status = False
                    r.save()
                    total_rubrosunemi += total_rubrosunemi
                    rubrospagados_noeliminados.append(
                        f'{r.persona.cedula} [{r.id}] porque el rubro no ha sido migrado a EPUNEMI')
                    print(
                        f"{contador}.- Se elimino rubro de unemi: [{r.persona.cedula} - {r.id}], no existe rubro en epunemi migrado.")
                cursor.close()
            print(f"Lista de rubros no eliminados: {rubrospagados_noeliminados}")
        except Exception as ex:
            transaction.set_rollback(True)
            line_eer = 'Error on line {}'.format(sys.exc_info()[-1].tb_lineno)
            print(f"Error: {ex.__str__()}. {line_eer}")


# eliminar_rubros_formacionejecutica()
def actualizar_nota_postulante(postulante, sel, valor):
    postulante = PersonaAplicarPartida.objects.get(pk=postulante.id)
    modeloevaluativo = postulante.partida.convocatoria.modeloevaluativoconvocatoria
    campomodelo = modeloevaluativo.campo(sel)
    try:
        if not valor:
            valor = null_to_decimal(float(valor), campomodelo.decimales)
        if valor >= campomodelo.notamaxima:
            valor = campomodelo.notamaxima
        elif valor <= campomodelo.notaminima:
            valor = campomodelo.notaminima
    except:
        valor = campomodelo.notaminima
    campo = postulante.campo(sel)
    campo.valor = valor
    campo.save()
    d = locals()
    exec(modeloevaluativo.logicamodelo, globals(), d)
    d['calculo_modelo_evaluativo'](postulante)
    postulante.nota_final = null_to_decimal(postulante.nota_final, 2)
    if postulante.nota_final > modeloevaluativo.notamaxima:
        postulante.nota_final = modeloevaluativo.notamaxima
    postulante.save()
    camposdependientes = []
    for campomodelo in modeloevaluativo.campos():
        if campomodelo.dependiente:
            camposdependientes.append(
                (campomodelo.htmlid(), postulante.valor_nombre_campo(campomodelo.nombre), campomodelo.decimales))

    postulante.actualiza_estado()
    postulante.actualiza_estadofinal()
    postulante.save()


def notas_postulate():
    try:
        ruta = os.path.join(SITE_ROOT, 'runback', 'arreglos', 'media')
        excel = openpyxl.load_workbook("{}{}result_notas_postulate_cmpl.xlsx".format(ruta, os.sep))
        hoja = excel['entrevista']
        numlist = hoja.rows
        cab = 0
        campo_pt = 'PT'
        campo_Erector = 'E1'
        campo_Evice = 'E2'
        campo_Eth = 'E3'
        lista_error = []
        for postu in numlist:
            if cab > 0:
                part_ = postu[1].value
                cedula_ = str(postu[15].value) if len(str(postu[15].value)) == 10 else '0' + str(postu[15].value)
                notas_pt = postu[7].value if postu[7].value != 'None' else 0
                notas_erector = postu[9].value if postu[9].value != 'None' else 0
                notas_evice = postu[10].value if postu[10].value != 'None' else 0
                notas_eth = postu[11].value if postu[11].value != 'None' else 0
                filtro = Q(cedula=cedula_)|Q(pasaporte=cedula_)
                persona = Persona.objects.get(filtro,status=True)
                partida = Partida.objects.get(status=True, id=part_)
                try:
                    aplica_partida = PersonaAplicarPartida.objects.get(status=True, persona=persona, partida=partida)
                except:
                    lista_error.append(f'No existe postulación de la persona {persona} con cedula ({cedula_}) en la convocatoria: {partida.convocatoria} - {partida}')
                    print(f'No existe postulación de la persona {persona} con cedula ({cedula_}) en la convocatoria: {partida.convocatoria} - {partida}')
                    continue
                modeloevaluativo = partida.convocatoria.modeloevaluativoconvocatoria
                if not modeloevaluativo:
                    lista_error.append(
                        f"Convocatoriasin modelo evaluativo: {partida.convocatoria}, persona ci: {cedula_}")
                try:
                    aplica_partida.valor_nombre_campo(campo_pt)
                except:
                    lista_error.append(f'Error al crear modelo evaluativo Persona: {cedula_}. Nota({campo_pt}): {notas_pt}')
                    print(f'Error al crear modelo evaluativo Persona: {cedula_}. Nota({campo_pt}): {notas_pt}')
                    continue
                try:
                    aplica_partida.valor_nombre_campo(campo_Erector)
                except:
                    lista_error.append(f'Error al crear modelo evaluativo Persona: {cedula_}. Nota({campo_Erector}): {notas_erector}')
                    print(f'Error al crear modelo evaluativo Persona: {cedula_}. Nota({campo_Erector}): {notas_erector}')
                    continue
                try:
                    aplica_partida.valor_nombre_campo(campo_Evice)
                except:
                    lista_error.append(f'Error al crear modelo evaluativo Persona: {cedula_}. Nota({campo_Evice}): {notas_evice}')
                    print(f'Error al crear modelo evaluativo Persona: {cedula_}. Nota({campo_Evice}): {notas_evice}')
                    continue
                try:
                    aplica_partida.valor_nombre_campo(campo_Eth)
                except:
                    lista_error.append(f'Error al crear modelo evaluativo Persona: {cedula_}. Nota({campo_Eth}): {notas_eth}')
                    print(f'Error al crear modelo evaluativo Persona: {cedula_}. Nota({campo_Eth}): {notas_eth}')
                    continue

                try:
                    actualizar_nota_postulante(aplica_partida, campo_pt, notas_pt)
                except:
                    lista_error.append(f'Error al guardar la nota Persona: {cedula_}. Nota({campo_pt}): {notas_pt}')
                    print(f'Error al guardar la nota Persona: {cedula_}. Nota({campo_pt}): {notas_pt}')
                    continue
                try:
                    actualizar_nota_postulante(aplica_partida, campo_Erector, notas_erector)
                except:
                    lista_error.append(f'Error al guardar la nota Persona: {cedula_}. Nota({campo_Erector}): {notas_erector}')
                    print(f'Error al guardar la nota Persona: {cedula_}. Nota({campo_Erector}): {notas_erector}')
                    continue
                try:
                    actualizar_nota_postulante(aplica_partida, campo_Evice, notas_evice)
                except:
                    lista_error.append(
                        f'Error al guardar la nota Persona: {cedula_}. Nota({campo_Evice}): {notas_evice}')
                    print(f'Error al guardar la nota Persona: {cedula_}. Nota({campo_Evice}): {notas_evice}')
                    continue
                try:
                    actualizar_nota_postulante(aplica_partida, campo_Eth, notas_eth)
                except:
                    lista_error.append(
                        f'Error al guardar la nota Persona: {cedula_}. Nota({campo_Eth}): {notas_eth}')
                    print(f'Error al guardar la nota Persona: {cedula_}. Nota({campo_Eth}): {notas_eth}')
                    continue
                print(f"No. {cab} {aplica_partida} - Notas: {notas_pt}/{notas_erector}/{notas_evice}/{notas_eth} - Total: {aplica_partida.nota_final}")
            cab += 1
        print(lista_error)
    except Exception as ex:
        print(f"Error en la linea: {sys.exc_info()[-1].tb_lineno}")
        print(ex)


# notas_postulate()

def migrar_departamentos_posgrado():
    with transaction.atomic():
        try:
            cont_p = 0
            departamentos_posgrado = DepPdip.objects.filter(status=True, id__in=[93])
            for deps in departamentos_posgrado:
                cont_p += 1
                if DepContrato.objects.values('id').filter(status=True, nombre=deps.nombre,
                                                           responsable=deps.responsable).exists():
                    new_dep = DepContrato.objects.filter(status=True, nombre=deps.nombre,
                                                         responsable=deps.responsable).order_by('-id').first()
                else:
                    new_dep = DepContrato(
                        nombre=deps.nombre,
                        responsable=deps.responsable,
                    )
                    new_dep.save()
                    new_dep.responsable_subrogante.set(deps.responsable_subrogante.all())
                    new_dep.save()
                print(f'{cont_p}.- Se migro el departamento {deps} al {new_dep}')
                cont_aux = 1
                for seccion in deps.secciondepartamento_set.filter(status=True):
                    if Gestion.objects.values('id').filter(status=True, departamento=new_dep,
                                                           gestion=seccion.descripcion, cargo=seccion.observacion,
                                                           responsable=seccion.responsable,
                                                           responsablesubrogante=seccion.responsablesubrogante).exists():
                        gestion = Gestion.objects.filter(status=True, departamento=new_dep, gestion=seccion.descripcion,
                                                         cargo=seccion.observacion, responsable=seccion.responsable,
                                                         responsablesubrogante=seccion.responsablesubrogante).order_by(
                            '-id').first()
                    else:
                        gestion = Gestion(
                            departamento=new_dep,
                            gestion=seccion.descripcion,
                            cargo=seccion.observacion,
                            responsable=seccion.responsable,
                            responsablesubrogante=seccion.responsablesubrogante
                        )
                        gestion.save()
                    print(f'----{cont_aux}.- Se migro la seccion {seccion} al {gestion}')
                    contratos = ContratoDip.objects.filter(status=True, seccion=seccion, gestion__isnull=True)
                    cont_aux2 = 1
                    for cont_ in contratos:
                        cont_.gestion = gestion
                        cont_.save()
                        cont_aux3 = 1
                        print(f'*******{cont_aux2}.- Se actualizo el contrato {cont_} con la gestion {gestion}')
                        cont_aux2 += 1
                        for areaprograma in cont_.contratoareaprograma_set.filter(status=True, gestion__isnull=True):
                            subsecciones = Gestion.objects.filter(status=True, departamento=new_dep,
                                                                  gestion=areaprograma.departamento.descripcion,
                                                                  cargo=areaprograma.departamento.observacion,
                                                                  responsable=areaprograma.departamento.responsable).order_by(
                                '-id').first()
                            areaprograma.gestion = subsecciones
                            areaprograma.save()
                            print(
                                f'>>>>>>{cont_aux3}.- Se actualizo el area o programa donde ejecuta sus actividades {areaprograma} con la gestion {subsecciones}')
                            cont_aux3 += 1
                    cont_aux += 1
            print('Fin de la migracion')
        except Exception as ex:
            transaction.set_rollback(True)
            print(f"Error en la linea: {sys.exc_info()[-1].tb_lineno}")
            print(ex)

# migrar_departamentos_posgrado()

def criterio_docente_recalcular():
    with transaction.atomic():
        try:
            cont = 1
            distributivo = DetalleDistributivo.objects.filter(status=True, criteriodocenciaperiodo_id=1021)
            print(len(distributivo))
            for dis in distributivo:
                print(f'{cont}.- Se borro el detalle distributivo {dis.__str__()}')
                cont += 1
                dis.delete()
        except Exception as ex:
            transaction.set_rollback(True)
            print(f"Error en la linea: {sys.exc_info()[-1].tb_lineno}")
            print(ex)
        finally:
            print('Proceso finalizado')


# criterio_docente_recalcular()

def actualizar_nombrehtlm_criterios_posgrado():
    with transaction.atomic():
        try:
            cont = 1
            criterio_orientacion_id = 7  # <CriterioDocencia: ORIENTACIÓN Y ACOMPAÑAMIENTO A TRAVÉS DE TUTORÍAS PRESENCIALES O VIRTUALES, INDIVIDUALES O GRUPALES>
            criterio_imparticion_clases_id = 46  # <CriterioDocencia: IMPARTICIÓN CLASES PRESENCIALES, VIRTUALES O EN LÍNEA, DE CARÁCTER TEÓRICO O PRÁCTICO, EN LA INSTITUCIÓN O FUERA DE ESTA, BAJO RESPONSABILIDAD Y DIRECCIÓN DE LA MISMA - EN LÍNEA>
            crierio_impartirclases_id = 118  # <CriterioDocencia: IMPARTIR CLASES>
            id_criterio_ayudantia = 19  # <CriterioDocencia: AYUDANTIA DE LABORATORIOS Y PRACTICAS DOCENTES - TÉCNICOS DOCENTES>
            id_criterio_class_nombramientoyprofesores = 16  # <CriterioDocencia: HORAS CLASE PROFESORES CON NOMBRAMIENTO Y PROFESORES DE CONTRATO>
            id_criterio_class_nombramientoyprofesores2 = 15  # <CriterioDocencia: HORAS CLASE PROFESORES CON NOMBRAMIENTO Y PROFESORES DE CONTRATO>
            id_criterio_hrs_prcacticas = 20  # <CriterioDocencia: HORAS DE PRÁCTICAS ASIGNATURA - SALUD>
            id_criterio_hrs_reemplazo = 27  # <CriterioDocencia: HORAS DE REEMPLAZO>
            id_criterio_imp_clases_prsenciales = 30  # <CriterioDocencia: IMPARTICIÓN CLASES PRESENCIALES, VIRTUALES O EN LÍNEA, DE CARÁCTER TEÓRICO O PRÁCTICO, EN LA INSTITUCIÓN O FUERA DE ESTA, BAJO RESPONSABILIDAD Y DIRECCIÓN DE LA MISMA.>
            ids_criterios_extra = [7, 15, 16, 19, 20, 27, 30, 46, 118]
            totales = 213
            totales_criterios = 140
            totales_excl_criterios = 73
            # criterios_extra = CriterioDocencia.objects.filter(status=True, id__in=ids_criterios_extra)
            criterios = CriterioDocenciaPeriodo.objects.filter(status=True, periodo__tipo__id__in=[3, 4])
            total_act = len(criterios)
            for crit in criterios:
                nombre_prev = crit.nombrehtml
                if crit.criterio.id == criterio_orientacion_id:
                    crit.nombrehtml = 'tutoriaacademica'
                elif crit.criterio.id == 118 or crit.criterio.id == 46 or crit.criterio.id == 30 or crit.criterio == 172:
                    crit.nombrehtml = 'impartirclase'
                crit.save()
                print(f'{cont}.- Actualizo el campo nombrehtml de {nombre_prev} a {crit.nombrehtml} del registro {crit.__str__()}. {cont}/{total_act}')
                cont += 1
        except Exception as ex:
            transaction.set_rollback(True)
            print(f"Error en la linea: {sys.exc_info()[-1].tb_lineno}")
            print(ex)
        finally:
            print('Proceso finalizado')


# actualizar_nombrehtlm_criterios_posgrado()


def migrar_rubros_congresos(id_congreso):
    try:
        print("**********INICIO DE MIGRAR RUBROS A EPUNEMI")
        inscritos = InscritoCongreso.objects.filter(status=True, requiere_certificado=True,congreso_id=id_congreso)
        fechamaxpago = datetime.strptime('24/12/2023',"%d/%m/%Y")
        rubros_totales_migrados = []
        rubros_sin_migrar = []
        cursor = connections['epunemi'].cursor()
        for ins in inscritos:
            print(f"****Inicia migrar rubros a: {ins.participante} ({ins.id})")
            rubro = Rubro.objects.filter(status=True, epunemi=True, tipo=ins.congreso.tiporubro,persona=ins.participante, congreso_id=id_congreso).order_by('-id').first()
            if not rubro:
                rubro = Rubro(tipo=ins.congreso.tipoubro,
                              persona=ins.participante,
                              congreso=ins.congreso,
                              relacionados=None,
                              nombre=ins.congreso.tipoubro.nombre + ' - ' + ins.congreso.nombre,
                              cuota=1,
                              fecha=datetime.now().date(),
                              fechavence=fechamaxpago,
                              valor=ins.tipoparticipacion.valor,
                              iva_id=1,
                              valoriva=0,
                              valortotal=ins.tipoparticipacion.valor,
                              saldo=ins.tipoparticipacion.valor,
                              epunemi=True,
                              observacion=ins.tipoparticipacion.nombre_completo(),
                              cancelado=False)
                rubro.save()
            rubro.fechavence = fechamaxpago
            rubro.save()
            hashed_password = make_password(ins.participante.cedula)
            sql = """SELECT pe.id,pe.usuario_id FROM sga_persona AS pe WHERE (pe.cedula='%s') AND pe.status=TRUE;  """ % (
                ins.participante.cedula)
            cursor.execute(sql)
            idpersona = cursor.fetchone()
            usuario_id = 'null'
            with transaction.atomic():
                try:
                    if idpersona is None:
                        if ins.participante.usuario:
                            sql = f"""SELECT us.id FROM auth_user us WHERE username = '{ins.participante.usuario.username}';
                                                                    """
                            cursor.execute(sql)
                            usuario = cursor.fetchone()
                            if usuario is None:
                                sql = f"""INSERT INTO auth_user (username, password, email, first_name, last_name, is_active, is_staff, is_superuser, date_joined)
                                            VALUES ('{ins.participante.usuario.username}', '{hashed_password}', '', '', '', TRUE, FALSE, FALSE, NOW());
                                                                            """
                                cursor.execute(sql)
                                sql = f"""
                                                                            SELECT us.id FROM auth_user us WHERE username = '{ins.participante.usuario.username}';
                                                                            """
                                cursor.execute(sql)
                                usuario_id = cursor.fetchone()
                                usuario_id = usuario_id[0]
                                print(f">>> Usuario creado en epunemi: {usuario_id}")
                            else:
                                usuario_id = usuario[0]
                                print(f">>> Usuario obtenido en epunemi: {usuario_id}")
                                sql = f"""UPDATE auth_user SET password = '{hashed_password}' WHERE id = {usuario_id}
                                """
                                cursor.execute(sql)
                        else:
                            sql = f"""SELECT us.id FROM auth_user us WHERE username = '{ins.participante.cedula}';
                                                                    """
                            cursor.execute(sql)
                            usuario = cursor.fetchone()
                            if usuario is None:
                                sql = f"""INSERT INTO auth_user (username, password, email, first_name, last_name, is_active, is_staff, is_superuser,date_joined)
                                        VALUES ('{ins.participante.cedula}', '{hashed_password}', '', '', '', TRUE, FALSE, FALSE, NOW());
                                                                            """
                                cursor.execute(sql)
                                sql = f"""SELECT us.id FROM auth_user us WHERE username = '{ins.participante.cedula}';
                                                                            """
                                cursor.execute(sql)
                                usuario_id = cursor.fetchone()
                                usuario_id = usuario_id[0]
                                print(f">>> Usuario creado en epunemi: {usuario_id}")
                            else:
                                usuario_id = usuario[0]
                                print(f">>> Usuario obtenido en epunemi: {usuario_id}")
                                sql = f"""UPDATE auth_user SET password = '{hashed_password}' WHERE id = {usuario_id}"""
                                cursor.execute(sql)
                        sql = f""" INSERT INTO sga_persona (status,usuario_id, nombres, apellido1, apellido2, cedula, ruc, pasaporte,
                            nacimiento, tipopersona, direccion,
                            telefono, email, contribuyenteespecial,
                            anioresidencia, nacionalidad, ciudad, referencia, emailinst, identificacioninstitucion,
                            regitrocertificacion, libretamilitar, servidorcarrera, concursomeritos, telefonoextension,
                            tipocelular, periodosabatico, real, lgtbi, datosactualizados, confirmarextensiontelefonia,
                            acumuladecimo, acumulafondoreserva, representantelegal, inscripcioncurso, unemi,
                            idunemi, sector,direccion2,num_direccion,telefono_conv)
                                    VALUES(TRUE,{usuario_id}, '{ins.participante.nombres}', '{ins.participante.apellido1}', '{ins.participante.apellido2}', '{ins.participante.cedula}', '{ins.participante.ruc}', '{ins.participante.pasaporte}', '/{ins.participante.nacimiento}/', {ins.participante.tipopersona if ins.participante.tipopersona else 1}, '{ins.participante.direccion}', '{ins.participante.telefono}', '{ins.participante.email}',
                                    FALSE, 0, '', '', '', '', '', '', '', FALSE, FALSE, '', 0, FALSE, TRUE, FALSE, 0, FALSE, TRUE, FALSE, FALSE,
                                    FALSE, FALSE, 0, '','','',''); """
                        cursor.execute(sql)
                        if ins.participante.sexo:
                            sql = """SELECT sexo.id FROM sga_sexo AS sexo WHERE sexo.id='%s'  AND sexo.status=TRUE;  """ % (
                                ins.participante.sexo.id)
                            cursor.execute(sql)
                            sexo = cursor.fetchone()
                            if sexo is not None:
                                sql = """UPDATE sga_persona SET sexo_id='%s' WHERE cedula='%s'; """ % (
                                    sexo[0], ins.participante.cedula)
                                cursor.execute(sql)
                        if ins.participante.pais:
                            sql = """SELECT pai.id FROM sga_pais AS pai WHERE pai.id='%s'  AND pai.status=TRUE;  """ % (
                                ins.participante.pais.id)
                            cursor.execute(sql)
                            pais = cursor.fetchone()

                            if pais is not None:
                                sql = """UPDATE sga_persona SET pais_id='%s' WHERE cedula='%s'; """ % (
                                    pais[0], ins.participante.cedula)
                                cursor.execute(sql)
                        sql = """SELECT pe.id FROM sga_persona AS pe WHERE (pe.cedula='%s' OR pe.pasaporte='%s' OR pe.ruc='%s') AND pe.status=TRUE;  """ % (
                            ins.participante.cedula,
                            ins.participante.cedula,
                            ins.participante.cedula)
                        cursor.execute(sql)
                        idpersona = cursor.fetchone()
                        alumnoepu = idpersona[0]

                        print(f">>> Persona creado en epunemi: {alumnoepu}")
                    else:
                        alumnoepu = idpersona[0]
                        print(f">>> Persona obtenida en epunemi: {alumnoepu}")
                    sql = """SELECT id FROM sagest_tipootrorubro WHERE idtipootrorubrounemi=%s; """ % (
                        rubro.tipo.id)
                    cursor.execute(sql)
                    registro = cursor.fetchone()
                    if registro is not None:
                        tipootrorubro = registro[0]
                    else:
                        sql = """SELECT id FROM sagest_centrocosto WHERE status=True AND unemi=True AND tipo=%s;""" % (
                            rubro.tipo.tiporubro)
                        cursor.execute(sql)
                        centrocosto = cursor.fetchone()
                        idcentrocosto = centrocosto[0]

                        # Consulto la cuenta contable
                        cuentacontable = CuentaContable.objects.get(partida=rubro.tipo.partida, status=True)
                        # Creo el tipo de rubro en epunemi
                        sql = """ Insert Into sagest_tipootrorubro (status, nombre, partida_id, valor, interface, activo, ivaaplicado_id, nofactura, exportabanco, cuentacontable_id, centrocosto_id, tiporubro, idtipootrorubrounemi, unemi, es_especie, es_convalidacionconocimiento)
                                                                                                                                                   VALUES(TRUE, '%s', %s, %s, FALSE, TRUE, %s, FALSE, TRUE, %s, %s, 1, %s, TRUE, FALSE, FALSE); """ % (
                            rubro.tipo.nombre, cuentacontable.partida.id, rubro.tipo.valor,
                            rubro.tipo.ivaaplicado.id, cuentacontable.id, idcentrocosto,
                            rubro.tipo.id)
                        cursor.execute(sql)

                        print(".:: Tipo de Rubro creado en EPUNEMI ::.")
                        # Obtengo el id recién creado del tipo de rubro
                        sql = """SELECT id FROM sagest_tipootrorubro WHERE idtipootrorubrounemi=%s; """ % (
                            rubro.tipo.id)
                        cursor.execute(sql)
                        registro = cursor.fetchone()
                        tipootrorubro = registro[0]
                    # pregunto si no existe rubro con ese id de unemi
                    sql = """SELECT id FROM sagest_rubro WHERE idrubrounemi=%s AND status=TRUE; """ % (
                        rubro.id)
                    cursor.execute(sql)
                    registrorubro = cursor.fetchone()
                    if registrorubro is None:
                        # Creo nuevo rubro en epunemi
                        sql = """ INSERT INTO sagest_rubro (status, persona_id, nombre, cuota, tipocuota, fecha, fechavence,
                                                                                                            valor, saldo, iva_id, valoriva, totalunemi, valortotal, cancelado, observacion, 
                                                                                                            idrubrounemi, tipo_id, fecha_creacion, usuario_creacion_id, tienenotacredito, valornotacredito, 
                                                                                                            valordescuento, anulado, compromisopago, refinanciado, bloqueado, bloqueadopornovedad, 
                                                                                                            titularcambiado, coactiva) 
                                                                                                          VALUES (TRUE, %s, '%s', %s, %s, '/%s/', '/%s/', %s, %s, %s, %s, %s, %s, %s, '%s', %s, %s, NOW(), 1, FALSE, 0, 0, FALSE, %s, %s, %s, FALSE, FALSE, %s); """ \
                              % (
                                  alumnoepu, rubro.nombre, rubro.cuota, rubro.tipocuota, rubro.fecha,
                                  rubro.fechavence, rubro.saldo,
                                  rubro.saldo, rubro.iva_id, rubro.valoriva, rubro.valor,
                                  rubro.valortotal, rubro.cancelado, rubro.observacion, rubro.id,
                                  tipootrorubro,
                                  rubro.compromisopago if rubro.compromisopago else 0,
                                  rubro.refinanciado, rubro.bloqueado, rubro.coactiva)
                        cursor.execute(sql)
                        sql = """SELECT id FROM sagest_rubro WHERE idrubrounemi=%s AND status=TRUE AND anulado=FALSE; """ % (
                            rubro.id)
                        cursor.execute(sql)
                        registro = cursor.fetchone()
                        rubroepunemi = registro[0]

                        rubro.idrubroepunemi = rubroepunemi
                        rubro.save()
                        print(f"**** Rubro creado en epunemi: {rubroepunemi}")
                        rubros_totales_migrados.append({'inscrito':ins.id,'rubro_unemi':rubro.id, 'rubro_epunemi':rubroepunemi})
                    else:
                        sql = """SELECT id FROM sagest_rubro WHERE idrubrounemi=%s AND status=TRUE AND cancelado=FALSE; """ % (
                            rubro.id)
                        cursor.execute(sql)
                        rubronoc = cursor.fetchone()
                        if rubronoc is not None:
                            sql = """SELECT id FROM sagest_pago WHERE rubro_id=%s; """ % (
                                registrorubro[0])
                            cursor.execute(sql)
                            tienerubropagos = cursor.fetchone()

                            if tienerubropagos is not None:
                                pass
                            else:
                                sql = """UPDATE sagest_rubro SET nombre = '%s', fecha = '/%s/', fechavence = '/%s/',
                                                                                                      valor = %s, saldo = %s, iva_id = %s, valoriva = %s, totalunemi = %s,
                                                                                                      valortotal = %s, observacion = '%s', tipo_id = %s
                                                                                                      WHERE id=%s; """ % (
                                    rubro.nombre, rubro.fecha, rubro.fechavence, rubro.saldo,
                                    rubro.saldo, rubro.iva_id,
                                    rubro.valoriva, rubro.valor, rubro.valortotal, rubro.observacion,
                                    tipootrorubro,
                                    registrorubro[0])
                                cursor.execute(sql)
                            rubro.idrubroepunemi = registrorubro[0]
                            rubro.save()
                            print(f"**** Rubro actualizado en epunemi: {registrorubro[0]}")
                            rubros_totales_migrados.append({'inscrito':ins.id,'rubro_unemi':rubro.id, 'rubro_epunemi':registrorubro[0]})
                except Exception as ex:
                    transaction.set_rollback(True)
                    print(f"Ocurrio un error al migrar el rubro {rubro.id} del inscrito {ins.participante}")
                    rubros_sin_migrar.append({'inscrito':ins.id, 'rubro_unemi':rubro.id})
        print(f"Total de rubros migrado: {len(rubros_totales_migrados)}")
        print(f"Total de rubros sin migrar: {len(rubros_sin_migrar)}")
        print(f"Rubros migrado: {rubros_totales_migrados}")
        print(f"Rubros sin migrar: {rubros_sin_migrar}")
    except Exception as ex:
        print(ex)
        print(sys.exc_info()[-1])
    finally:
        print("*****FIN DEL PROCESO********")

migrar_rubros_congresos(20)