# -*- coding: latin-1 -*-
import json
import sys
import io
import xlsxwriter
import openpyxl
from django.template.loader import get_template
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models.query_utils import Q
from django.http import HttpResponseRedirect, HttpResponse, JsonResponse
from django.shortcuts import render, redirect
from decorators import secure_module, last_access
from sga.commonviews import adduserdata
from sga.forms import InstitucionEducacionSuperiorForm
from sga.funciones import MiPaginador, log, elimina_tildes
from sga.models import InstitucionEducacionSuperior, Pais


@login_required(redirect_field_name='ret', login_url='/loginsga')
@secure_module
@last_access
@transaction.atomic()
def view(request):
    global ex
    data = {}
    adduserdata(request, data)
    persona = request.session['persona']
    if request.method == 'POST':
        action = request.POST['action']
        if action == 'ImportUniversidad':
            try:
                with transaction.atomic():
                    excel = request.FILES['excel']
                    wb = openpyxl.load_workbook(excel)
                    worksheet = wb.worksheets[0]
                    count = 0
                    counter = 0
                    linea = 1
                    for row in worksheet.iter_rows():
                        currentValues = [str(cell.value) for cell in row]
                        if linea >= 2:
                            if currentValues[0] == 'None':
                                messages.error(request, 'REVISAR LINEA [{}],  FORMATO FILAS SIN NOMBRE DE PAIS, VERIFIQUE QUE EL ARCHIVO NO TENGA COLUMNAS VACIAS AL FINAL.'.format(linea))
                                transaction.set_rollback(True)
                                return redirect(request.path)
                            if currentValues[1] == 'None':
                                messages.error(request, 'REVISAR LINEA [{}],  FORMATO FILAS SIN NOMBRE DE UNIVERSIDAD, VERIFIQUE QUE EL ARCHIVO NO TENGA COLUMNAS VACIAS AL FINAL.'.format(linea))
                                transaction.set_rollback(True)
                                return redirect(request.path)
                            if currentValues[0] == '':
                                messages.error(request, 'REVISAR LINEA [{}],  LA COLUMNA DE PAIS DEBE ESTAR LLENA.'.format(linea))
                                transaction.set_rollback(True)
                                return redirect(request.path)
                            if currentValues[1] == '':
                                messages.error(request, 'REVISAR LINEA [{}],  LA COLUMNA DE UNIVERSIDAD DEBE ESTAR LLENA.'.format(linea))
                                transaction.set_rollback(True)
                                return redirect(request.path)
                            pais = currentValues[0].replace('\t','').rstrip() if currentValues[0] != 'None' else None
                            universidad = currentValues[1].replace('\t','').rstrip() if currentValues[1] != 'None' else None
                            if Pais.objects.filter(status=True, nombre__icontains=pais).exists():
                                if universidad:
                                    pais = Pais.objects.filter(status=True, nombre__icontains=pais).first()
                                    universidad = elimina_tildes(universidad.upper())
                                    if InstitucionEducacionSuperior.objects.filter(nombre__unaccent__icontains=universidad).exists():
                                        count += 1
                                        car = InstitucionEducacionSuperior.objects.filter(status=True, nombre__unaccent__icontains=universidad).first()
                                        car.pais = pais
                                        car.save(request)
                                        log(u'Adiciono Edito: %s' % car, request, "change")
                                        return JsonResponse({"result": False, "message": "Registro Exitoso"})
                                    else:
                                        car = InstitucionEducacionSuperior(nombre=universidad,
                                                                           pais=pais)
                                        car.save(request)
                                        log(u'Adiciono Universidad: %s' % car, request, "ImportUniversidad")
                                        counter += 1
                                else:
                                    messages.error(request, 'REVISAR LINEA [{}],  LA COLUMNA DE UNIVERSIDAD DEBE ESTAR LLENA.'.format(linea))
                                    transaction.set_rollback(True)
                                    return redirect(request.path)
                            else:
                                messages.error(request, '{} NO SE ENCUENTRA REGISTRADO EN TABLA PAISES.'.format(pais))
                                messages.info(request, 'VERIFICAR SI EL PAIS A INGRESAR EXISTE O NO TENGA UN ERROR DE ESCRITURA')
                                transaction.set_rollback(True)
                                return redirect(request.path)
                        linea += 1
                    if count > 0:
                        messages.info(request, 'Se actualizaron {} registros ya existentes.'.format(str(count)))
                    if counter > 0:
                        messages.success(request, 'Se agregaron {} universidades'.format(counter))
            except Exception as ex:
                transaction.set_rollback(True)
                print(ex)
                print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
                messages.error(request, ex)
                messages.error(request, 'Acaba de ocurrir un error al intertar cargar la linea {}, Acción fue cancelada.'.format(linea))
            return redirect('{}'.format(request.path))

        elif action == 'add':
            try:
                f = InstitucionEducacionSuperiorForm(request.POST)
                if f.is_valid():
                    institucioneducacionsuperior = InstitucionEducacionSuperior(nombre=f.cleaned_data['nombre'],
                                                                                pais=f.cleaned_data['pais'],
                                                                                codigo=f.cleaned_data['codigo'])
                    institucioneducacionsuperior.save(request)
                    log(u'Adiciono carrera: %s' % institucioneducacionsuperior, request, "add")
                    return JsonResponse({"result": "ok", "id": institucioneducacionsuperior.id})
                else:
                     raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})
        elif action == 'addUni':
            try:
                f = InstitucionEducacionSuperiorForm(request.POST)
                if f.is_valid():
                    institucioneducacionsuperior = InstitucionEducacionSuperior(nombre=f.cleaned_data['nombre'],
                                                                                pais=f.cleaned_data['pais'],
                                                                                codigo=f.cleaned_data['codigo'])
                    institucioneducacionsuperior.save(request)
                    log(u'Adiciono carrera: %s' % institucioneducacionsuperior, request, "addUni")
                    return JsonResponse({"result": False, "message": "Registro Exitoso"})
                else:
                     raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'edit':
            try:
                institucioneducacionsuperior = InstitucionEducacionSuperior.objects.get(pk=request.POST['id'])
                f = InstitucionEducacionSuperiorForm(request.POST)
                if f.is_valid():
                    institucioneducacionsuperior.nombre = f.cleaned_data['nombre']
                    institucioneducacionsuperior.pais = f.cleaned_data['pais']
                    institucioneducacionsuperior.codigo = f.cleaned_data['codigo']
                    institucioneducacionsuperior.save(request)
                    log(u'Modifico carrera: %s' % institucioneducacionsuperior, request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                    pass
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        return JsonResponse({"result": "bad", "mensaje": u"Solicitud Incorrecta."})
    else:
        data['title'] = u'Institucion Educación Superior'
        if 'action' in request.GET:
            action = request.GET['action']
            if action == 'ImportUniversidad':
                try:
                    template = get_template("institucioneducacionsuperior/modal/excel.html")
                    return JsonResponse({'result': True, 'data': template.render(data)})
                except Exception as ex:
                    pass
            elif action == 'excelprograma':
                    try:
                        __author__ = 'Unemi'

                        output = io.BytesIO()
                        workbook = xlsxwriter.Workbook(output)
                        ws = workbook.add_worksheet('exp_xls_post_part')
                        ws.set_column(0, 0, 10)
                        ws.set_column(1, 1, 40)
                        ws.set_column(2, 2, 50)
                        ws.set_column(3, 3, 20)
                        ws.set_column(4, 4, 20)
                        ws.set_column(5, 5, 20)
                        ws.set_column(6, 6, 40)
                        ws.set_column(7, 7, 35)
                        ws.set_column(8, 8, 40)
                        ws.set_column(9, 9, 20)
                        ws.set_column(10, 10, 20)

                        #                   ws.columm_dimensions['A'].width = 20

                        # formatotitulo = workbook.add_format(
                        #     {'bold': 1, 'text_wrap': True, 'border': 1, 'align': 'center', 'valign': 'middle',
                        #      'fg_color': '#A2D0EC'})
                        formatotitulo_filtros = workbook.add_format(
                            {'bold': 1, 'text_wrap': True, 'border': 1, 'fg_color': '#EBF5FB'})

                        formatoceldacab = workbook.add_format(
                            {'align': 'center', 'bold': 1, 'border': 1, 'text_wrap': True, 'fg_color': '#0b2f44', 'font_color':'white'})
                        formatoceldaleft = workbook.add_format(
                            {'text_wrap': True, 'align': 'center', 'valign': 'vcenter', 'border': 1})

                        ws.write(0, 0, 'ID', formatoceldacab)
                        ws.write(0, 1, 'NOMBRE', formatoceldacab)
                        ws.write(0, 2, 'CODIGO', formatoceldacab)
                        ws.write(0, 3, 'PAIS', formatoceldacab)

                        listaprogramas = InstitucionEducacionSuperior.objects.filter(status=True).order_by('-id')

                        filas_recorridas = 2
                        for programa in listaprogramas:

                            ws.write('A%s' % filas_recorridas, str(programa.id), formatoceldaleft)
                            ws.write('B%s' % filas_recorridas, str(programa.nombre), formatoceldaleft)
                            ws.write('C%s' % filas_recorridas, str(programa.codigo), formatoceldaleft)
                            ws.write('D%s' % filas_recorridas, str(programa.pais), formatoceldaleft)

                            filas_recorridas += 1

                        workbook.close()
                        output.seek(0)
                        filename = 'Matriz_Programas.xlsx'
                        response = HttpResponse(output,
                                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                        response['Content-Disposition'] = 'attachment; filename=%s' % filename
                        return response
                    except Exception as ex:
                       pass

            elif action == 'add':
                    try:
                        data['title'] = u'Nueva Institucion Educación Superior'
                        form = InstitucionEducacionSuperiorForm()
                        data['form'] = form
                        return render(request, "institucioneducacionsuperior/add.html", data)
                    except Exception as ex:
                        pass
            elif action == 'addUni':
                    try:
                        data['title'] = u'Nueva Institucion Educación Superior'
                        form = InstitucionEducacionSuperiorForm()
                        data['form'] = form
                        template = get_template("institucioneducacionsuperior/modal/ModalAgregarUniversidad.html")
                        return JsonResponse({'result': True, 'data': template.render(data)})
                    except Exception as ex:
                        pass
            elif action == 'import':
                try:
                    data['title'] = u'Importar Unidades Educativas'
                    form = InstitucionEducacionSuperiorForm
                    data['form'] = form
                    return render(request, "institucioneducacionsuperior/importExcel.html", data)
                except Exception as ex:
                    pass

            elif action == 'edit':
                try:
                    data['title'] = u'Editar Institución Educación Superior'
                    data['institucioneducacionsuperior'] = institucioneducacionsuperior = InstitucionEducacionSuperior.objects.get(pk=request.GET['id'])
                    form = InstitucionEducacionSuperiorForm(initial={'nombre': institucioneducacionsuperior.nombre,
                                                                     'pais': institucioneducacionsuperior.pais,
                                                                     'codigo': institucioneducacionsuperior.codigo})
                    data['form'] = form
                    return render(request, "institucioneducacionsuperior/edit.html", data)
                except Exception as ex:
                    pass

            return HttpResponseRedirect(request.path)
        else:
            search = None
            ids = None
            if 'id' in request.GET:
                ids = request.GET['id']
                institucioneducacionsuperior = InstitucionEducacionSuperior.objects.filter(id=ids)
            elif 's' in request.GET:
                search = request.GET['s']
                institucioneducacionsuperior = InstitucionEducacionSuperior.objects.filter(Q(nombre__icontains=search) | Q(codigo__icontains=search)).distinct()
            else:
                institucioneducacionsuperior = InstitucionEducacionSuperior.objects.all()
            paging = MiPaginador(institucioneducacionsuperior, 25)
            p = 1
            try:
                paginasesion = 1
                if 'paginador' in request.session:
                    paginasesion = int(request.session['paginador'])
                if 'page' in request.GET:
                    p = int(request.GET['page'])
                else:
                    p = paginasesion
                try:
                    page = paging.page(p)
                except:
                    p = 1
                page = paging.page(p)
            except:
                page = paging.page(p)
            request.session['paginador'] = p
            data['paging'] = paging
            data['rangospaging'] = paging.rangos_paginado(p)
            data['page'] = page
            data['search'] = search if search else ""
            data['ids'] = ids if ids else ""
            data['institucioneducacionsuperiors'] = page.object_list
            return render(request, "institucioneducacionsuperior/view.html", data)