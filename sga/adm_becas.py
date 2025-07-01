# -*- coding: UTF-8 -*-
import json
import os
import sys
import threading
import pyqrcode
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.exceptions import ObjectDoesNotExist
from django.db import connection, connections
from django.db import transaction
from django.db.models import Count, PROTECT, Sum, Avg, Min, Max
from django.db.models.query_utils import Q
from django.http import HttpResponseRedirect, JsonResponse, HttpResponse
from django.shortcuts import render
import xlwt
from django.template.loader import get_template
from django.template import Context
from django.forms.models import model_to_dict
from django.core.files import File
from openpyxl import load_workbook
import xlsxwriter
from webpush import send_user_notification
from xlwt import *
import random
import time
from datetime import datetime, timedelta
from decimal import Decimal

from django.db.models import Max
import xlrd as xlrd

from decorators import secure_module, last_access
from sagest.commonviews import secuencia_convenio_devengacion
from sagest.models import DistributivoPersona
from settings import ARCHIVO_TIPO_GENERAL, DEBUG, SITE_STORAGE, MEDIA_ROOT, MEDIA_URL
from sga.commonviews import adduserdata, secuencia_contrato_beca, secuencia_solicitud_pago_beca, obtener_reporte, traerNotificaciones
from sga.excelbackground import reportebecadosmatrizsiiescaces
from sga.forms import BecaTipoForm, BecaUtilizacionForm, BecaRequisitosForm, BecaAprobarArchivoForm, \
    BecaAprobarArchivoUtilizacionForm, BecaPeriodoForm, BecaAsignacionForm, ImportarBecaForm, BecaAsignacionManualForm, \
    BecaSolicitudAnulacionForm, BecaAsignacion2Form, BecaSolicitudRechazaOnLineForm, BecaValidarCedulaForm, \
    BecaSolicitudSubirCedulaForm, BecaAsignacionSubirEvidenciaForm, RepresentanteSolidarioForm, \
    ImportarArchivoPagoBecaXLSForm, BecaComprobanteVentaEditForm, BecaComprobanteVentaValidaForm, \
    BecaComprobanteEliminarForm, BecaTipoConfiguracionForm, DetalleRequisitoBecaForm, ImportarPreinscritosBecaForm, \
    SubirArchivoRequisitoBecaForm, ImportarArchivoBecariosForm, DocumentoBecaTipoConfiguracionForm, FiltroBecadoValidacionForm, AdicionarPreBecadoForm
from sga.funciones import MiPaginador, log, variable_valor, convertir_fecha, null_to_numeric, convertir_hora, \
    generar_nombre, convertir_fecha_invertida, convertir_fecha_invertida_hora, convertir_fecha_hora, \
    convertir_fecha_hora_invertida, fechaformatostr, fechaletra_corta, validarcedula, cuenta_email_disponible, \
    MiPaginador, log, variable_valor, null_to_decimal, lista_mejores_promedio_beca_v2, lista_deportista_beca_v2, \
    lista_gruposocioeconomico_beca_v2, lista_mejores_promedio_beca, lista_discapacitado_beca, lista_deportista_beca, \
    lista_migrante_exterior_beca, lista_etnia_beca, lista_gruposocioeconomico_beca, \
    leer_fecha_hora_letra_formato_fecha_hora, ok_json, bad_json, listado_becados_por_desviacionestandar_todas_carrera, \
    listado_incripciones_reconocimiento_academico, listado_becados_por_desviacionestandar_todas_carreranew, \
    listado_becados_por_desviacionestandar_todas_carreramejorado, asignar_orden_portipo_beca, lista_mejores_promedio_beca_v3, \
    remover_caracteres_tildes_unicode
from sga.funciones_becas import generar_precandidatos_becas, generar_reportes_por_query
from sga.funcionesxhtml2pdf import conviert_html_to_pdf, conviert_html_to_pdfsaveqr_generico
from sga.graduados import fecha_letra
from sga.models import Inscripcion, BecaTipo, \
    BecaAsignacion, BecaUtilizacion, BecaRequisitos, MatriculaGrupoSocioEconomico, Matricula, Periodo, MateriaAsignada, \
    BecaSolicitud, BecaDetalleSolicitud, BecaDetalleUtilizacion, BecaPeriodo, CUENTAS_CORREOS, TIPO_NUE_RENO, \
    MESES_CHOICES, Archivo, Persona, Carrera, Modalidad, BecaTablaAsistencia, ESTADO_BECA_ASIG, BecaSolicitudRecorrido, \
    miinstitucion, ConfirmaCapacidadTecnologica, SolicitudPagoBeca, SolicitudPagoBecaDetalle, ArchivoPagoBeca, \
    BecaComprobanteRevision, BecaComprobanteVenta, Coordinacion, Materia, Malla, AsignaturaMalla, RecordAcademico, \
    NivelMalla, TarjetaRegistroAcademico, PreInscripcionBeca, Discapacidad, Raza, Pais, InscripcionNivel, \
    BecaTipoConfiguracion, ESTADO_ACEPTACION_BECA, Reporte, Notificacion, DetalleRequisitoBeca, FUNCIONES_REQUISITOSBECAS_EJECUTAR, \
    PreInscripcionBecaRequisito, HistorialPreInscripcionBecaRequisito, PreInscripcionBecaRequisitoArchivo, ESTADO_MATRICULA, ArchivoBecarios, RequisitoBeca,\
    DocumentoBecaTipoConfiguracion, FUNCIONES_DOCUMENTOSBECAS_EJECUTAR,CuentaBancariaPersona
from sagest.models import Banco,TipoCuentaBanco
from socioecon.models import GrupoSocioEconomico
from sga.tasks import send_html_mail, conectar_cuenta
from dateutil.relativedelta import relativedelta
from django.db.models.functions import ExtractYear

from sga.templatetags.sga_extras import encrypt
from sga.reportes import elimina_tildes, transform_jasperstarter_new, run_report_v1
from utils.filtros_genericos import filtro_persona_select_v2

@login_required(redirect_field_name='ret', login_url='/loginsga')
@secure_module
@last_access
@transaction.atomic()
def view(request):
    data = {}
    adduserdata(request, data)
    data['persona'] = persona = request.session['persona']
    data['periodo'] = periodosesion = request.session['periodo']
    periodo = Periodo.objects.get( id=periodosesion.id)
    if not periodo.tipo_id==2:
        return HttpResponseRedirect("/?info=El periodo seleccionado no aplica a Becas.")
    data['periodosregular'] = Periodo.objects.filter(status=True, tipo__id=2, activo=True).order_by('-inicio')
    data['modalidadesinscripcion'] = Modalidad.objects.filter(status=True)

    if BecaRequisitos.objects.values('id').filter(status=True, vigente=True).exists():
        data['requisitos'] = BecaRequisitos.objects.filter(status=True, vigente=True).order_by('id')
        requisito = BecaRequisitos.objects.filter(status=True, vigente=True).order_by('id')[0]
    y = Periodo.objects.filter(status=True, id__lte=periodo.id, tipo=2, aplicabeca=True).exclude(
        id=periodo.id).order_by("-inicio")
    if not y[0].aplicabeca:
        return HttpResponseRedirect("/?info=El periodo anterior no aplica a Becas, seleccione otro.")
    anterior = y[0]
    data['anterior'] = anterior.id
    data['periodoanterior'] = anterior
    data['fechafinperiodonterior'] = fechafinperiodonterior = anterior.fin
    data['idper'] = periodo.id
    miscarreras = persona.mis_carreras()
    if BecaPeriodo.objects.values('id').filter(status=True, periodo=periodo).exists():
        data['becaperiodo'] = BecaPeriodo.objects.filter(status=True, periodo=periodo)[0]
    if request.method == 'POST':
        action = request.POST['action']

        if action == 'listapagadoscorreo':
            try:
                if not SolicitudPagoBecaDetalle.objects.values('id').filter(asignacion__solicitud__periodo=periodo,
                                                                            pagado=True,
                                                                            notificadopago=False
                                                                            ).exists():
                    return JsonResponse({"result": "bad",
                                         "mensaje": u"No existen becas pagadas pendientes de notificar por e-mail para el periodo seleccionado."})

                pagados = SolicitudPagoBecaDetalle.objects.values('id').filter(asignacion__solicitud__periodo=periodo,
                                                                               pagado=True,
                                                                               notificadopago=False
                                                                               )
                becadospagados = []
                for pagado in pagados:
                    beca = {'id': pagado['id']}
                    becadospagados.append(beca)
                return JsonResponse({"result": "ok", "cantidad": len(becadospagados), "becadospagados": becadospagados})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al generar la consulta."})

        elif action == 'enviaremailpagado':
            try:
                import time
                pago = SolicitudPagoBecaDetalle.objects.get(pk=int(request.POST['idpago']))
                pago.notificadopago = True
                pago.save(request)
                cuenta = cuenta_email_disponible()

                # Envio de e-mail de notificación al solicitante
                asunto = "Beca estudiantil"

                send_html_mail(asunto,
                               "emails/notificarpagobeca.html",
                               {    'pago': pago,
                                    'sistema': request.session['nombresistema'],
                                },
                               pago.asignacion.solicitud.inscripcion.persona.lista_emails_envio(),
                               [],
                               cuenta=CUENTAS_CORREOS[cuenta][1]
                               )

                # Temporizador para evitar que se bloquee el servicio de gmail
                time.sleep(3)
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al enviar el e-mail."})

        elif action == 'listaacreditadoscorreo':
            try:
                if not SolicitudPagoBecaDetalle.objects.values('id').filter(asignacion__solicitud__periodo=periodo,
                                                                            pagado=True,
                                                                            acreditado=True,
                                                                            notificadopago=True,
                                                                            notificadoacredita=False
                                                                            ).exists():
                    return JsonResponse({"result": "bad",
                                         "mensaje": u"No existen becas acreditadas pendientes de notificar por e-mail para el periodo seleccionado."})

                pagados = SolicitudPagoBecaDetalle.objects.values('id').filter(asignacion__solicitud__periodo=periodo,
                                                                               pagado=True,
                                                                               acreditado=True,
                                                                               notificadopago=True,
                                                                               notificadoacredita=False
                                                                               )
                becadosacreditados = []
                for pagado in pagados:
                    beca = {'id': pagado['id']}
                    becadosacreditados.append(beca)
                return JsonResponse(
                    {"result": "ok", "cantidad": len(becadosacreditados), "becadosacreditados": becadosacreditados})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al generar la consulta."})

        elif action == 'enviaremailacreditado':
            try:
                import time
                pago = SolicitudPagoBecaDetalle.objects.get(pk=int(request.POST['idpago']))
                pago.notificadoacredita = True
                pago.save(request)
                cuenta = cuenta_email_disponible()

                # Envio de e-mail de notificación al solicitante
                tituloemail = "Proceso de Adjudicación por Beca - Acreditado por Ministerio de Economía y Finanzas"

                send_html_mail(tituloemail,
                               "emails/notificaracreditadobeca.html",
                               {'sistema': u'SGA - UNEMI',
                                'fecha': datetime.now().date(),
                                'hora': datetime.now().time(),
                                'saludo': 'Estimada' if pago.asignacion.solicitud.inscripcion.persona.sexo_id == 1 else 'Estimado',
                                'estudiante': pago.asignacion.solicitud.inscripcion.persona.nombre_completo_minus(),
                                'autoridad2': '',
                                'fechaacredita': pago.fechaacredita,
                                't': miinstitucion()
                                },
                               pago.asignacion.solicitud.inscripcion.persona.lista_emails_envio(),
                               [],
                               cuenta=CUENTAS_CORREOS[cuenta][1]
                               )

                # Temporizador para evitar que se bloquee el servicio de gmail
                time.sleep(5)

                # cuenta = 22
                # Envío de e-mail indicando el plazo para la subida del comprobante de venta
                # tituloemail = "Comunicado Dirección de Bienestar Universitario"

                # send_html_mail(tituloemail,
                #                "emails/notificarsubirfactura.html",
                #                {'sistema': u'SGA - UNEMI',
                #                 'fecha': datetime.now().date(),
                #                 'hora': datetime.now().time(),
                #                 'saludo': 'Estimada' if pago.asignacion.solicitud.inscripcion.persona.sexo_id == 1 else 'Estimado',
                #                 'estudiante': pago.asignacion.solicitud.inscripcion.persona.nombre_completo_inverso(),
                #                 'autoridad2': '',
                #                 'fechapago': pago.fechapago,
                #                 't': miinstitucion()
                #                 },
                #                pago.asignacion.solicitud.inscripcion.persona.lista_emails_envio(),
                #                [],
                #                cuenta=CUENTAS_CORREOS[cuenta][1]
                #                )

                # Temporizador para evitar que se bloquee el servicion de gmail
                # time.sleep(5)

                return JsonResponse({"result": "ok"})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al enviar el e-mail."})

        elif action == 'generanumerocontrato':
            try:
                beca = BecaAsignacion.objects.get(pk=int(request.POST['id']))
                if not beca.numerocontrato:
                    secuencia = secuencia_contrato_beca(beca.solicitud.becatipo.id)

                    if BecaAsignacion.objects.filter(numerocontrato=secuencia, status=True).exists():
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"Error al generar el contrato, intente nuevamente"})

                    beca.fechacontrato = datetime.now().date()
                    beca.numerocontrato = secuencia
                    beca.save(request)
                    log(u'Editó registro de beca: %s' % beca, request, "edit")

                return JsonResponse({"result": "ok"})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al generar la secuencia del contrato."})

        elif action == 'actualizafechaacta':
            try:
                beca = BecaAsignacion.objects.get(pk=int(request.POST['id']))

                necesidad = beca.solicitud.becasolicitudnecesidad_set.all()[0]
                if necesidad.necesidad == 1:
                    if not necesidad.serietablet or not necesidad.imeitablet or not necesidad.numerosimcard:
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"Los campos serie, imei y número de simcard están vacíos."})
                else:
                    if not necesidad.numerosimcard:
                        return JsonResponse({"result": "bad", "mensaje": u"El campo número de simcard está vacío."})

                if not beca.fechaacta:
                    beca.fechaacta = datetime.now().date()
                    beca.numeroacta = beca.numerocontrato
                    beca.save(request)
                    log(u'Editó fecha del acta de registro de beca: %s' % beca, request, "edit")

                return JsonResponse({"result": "ok"})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al generar el acta."})

        elif action == 'contratobecapdf':
            try:
                data = {}

                beca = BecaAsignacion.objects.get(pk=int(request.POST['id']))

                # data['fechainiciocap'] = str(capacitacion.fechainicio.day) + " de " + \
                #                          MESES_CHOICES[capacitacion.fechainicio.month - 1][
                #                              1].capitalize() + " del " + str(capacitacion.fechainicio.year)
                # data['fechafincap'] = str(capacitacion.fechafin.day) + " de " + \
                #                       MESES_CHOICES[capacitacion.fechafin.month - 1][1].capitalize() + " del " + str(
                #     capacitacion.fechafin.year)

                data['numerocontrato'] = "CONTRATO Nº 1S-2020-" + str(beca.numerocontrato).zfill(4)
                data['beca'] = beca
                data['fechacontrato'] = fechaletra_corta(beca.fechacontrato)

                return conviert_html_to_pdf(
                    'adm_becas/contratobecapdf.html',
                    {
                        'pagesize': 'A4',
                        'data': data,
                    }
                )
            except Exception as ex:
                return HttpResponseRedirect(
                    "/adm_becas?info=%s" % "Error al generar el reporte del contrato de la beca")

        elif action == 'actaentregarecepcionpdf':
            try:
                data = {}

                beca = BecaAsignacion.objects.get(pk=int(request.POST['id']))
                umat = Matricula.objects.get(inscripcion=beca.solicitud.inscripcion, nivel__periodo_id=110, status=True,
                                             estado_matricula__in=[2, 3])
                facultad = str(umat.nivel.coordinacion())
                carrera = str(umat.inscripcion.carrera)
                modalidad = str(umat.inscripcion.modalidad)
                items = []
                necesidad = beca.solicitud.becasolicitudnecesidad_set.all()[0]
                if necesidad.necesidad == 1:
                    items.append(['TABLET' + ' SERIE: ' + necesidad.serietablet + ', IMEI: ' + necesidad.imeitablet])
                    items.append(['SIMCARD' + ' # ' + necesidad.numerosimcard])
                else:
                    items.append(['SIMCARD' + ' # ' + necesidad.numerosimcard])

                data['numeroacta'] = "Nº " + str(beca.numeroacta).zfill(4)
                data['beca'] = beca
                data['fechaacta'] = fechaletra_corta(beca.fechaacta)
                data['facultad'] = facultad
                data['carrera'] = carrera
                data['modalidad'] = modalidad
                data['items'] = items

                return conviert_html_to_pdf(
                    'adm_becas/actaentregarecepcionpdf.html',
                    {
                        'pagesize': 'A4',
                        'data': data,
                    }
                )
            except Exception as ex:
                return HttpResponseRedirect(
                    "/adm_becas?info=%s" % "Error al generar el reporte del acta entrega-recepción de la beca")

        elif action == 'reportesolicitudpagopdf':
            try:
                data = {}

                solicitudpago = SolicitudPagoBeca.objects.get(pk=int(request.POST['id']))
                detallesolicitud = solicitudpago.solicitudpagobecadetalle_set.all().order_by(
                    'asignacion__solicitud__inscripcion__persona__apellido1',
                    'asignacion__solicitud__inscripcion__persona__apellido2',
                    'asignacion__solicitud__inscripcion__persona__nombres')
                # detallesolicitud = detallesolicitud[:500]
                data['fechasolicitud'] = str(solicitudpago.fecha.day) + " de " + \
                                         MESES_CHOICES[solicitudpago.fecha.month - 1][1].capitalize() + " del " + str(
                    solicitudpago.fecha.year)
                data['numeroreporte'] = "Nº " + str(solicitudpago.numerosolicitud).zfill(5)
                data['solicitudpago'] = solicitudpago
                data['detallesolicitud'] = detallesolicitud
                data['periodo'] = solicitudpago.periodo.nombre
                cantidadmeses = BecaAsignacion.objects.filter(
                    pk__in=detallesolicitud.values_list('asignacion_id', flat=True)).aggregate(
                    cantidadmeses=Max("cantidadmeses"))['cantidadmeses']
                data['totalcantidadmeses'] = cantidadmeses
                contador_letra = ['1er MES', '2do MES', '3er MES', '4to MES', '5to MES', '6to MES', '7mo MES',
                                  '8vo MES', '9no MES']
                data['lista_meses'] = contador_letra[:cantidadmeses]
                data['persona'] = persona
                return conviert_html_to_pdf(
                    'adm_becas/solicitudpagopdf.html',
                    {
                        'pagesize': 'A4',
                        'data': data,
                    }
                )
            except Exception as ex:
                return HttpResponseRedirect("/adm_becas?info=%s" % "Error al generar el reporte de solicitud de pago")

        elif action == 'reporteseguimientopdf':
            try:
                data = {}

                data['periodo'] = periodo.nombre

                solicitudes = BecaSolicitud.objects.filter(status=True, periodo=periodo, periodocalifica__isnull=False,
                                                           estado=2)
                totalaprobada = solicitudes.count()
                totalbecaaceptada = solicitudes.filter(estado=2, becaaceptada=2).count()
                totalaceptada1vez = solicitudes.filter(estado=2, becaaceptada=2, tiposolicitud=1).count()
                totalaceptadarenovacion = solicitudes.filter(estado=2, becaaceptada=2, tiposolicitud=2).count()
                totalbecarechazada = solicitudes.filter(estado=2, becaaceptada=3).count()
                totalbecapendiente = totalaprobada - (totalbecaaceptada + totalbecarechazada)

                becas = []
                becas.append(['TOTAL ACEPTADAS POR ESTUDIANTES', totalbecaaceptada])
                becas.append(['TOTAL RECHAZADAS POR ESTUDIANTES', totalbecarechazada])
                becas.append(['TOTAL PENDIENTES DE ACEPTAR/RECHAZAR POR ESTUDIANTES', totalbecapendiente])
                data['becas'] = becas
                data['totalbecas'] = totalaprobada

                aceptadas = []
                aceptadas.append(['TOTAL PRIMERA VEZ', totalaceptada1vez])
                aceptadas.append(['TOTAL RENOVACIÓN', totalaceptadarenovacion])
                data['becasaceptadas'] = aceptadas
                data['totalbecasaceptadas'] = totalbecaaceptada

                asignados = BecaAsignacion.objects.filter(status=True, solicitud__periodo=periodo)
                totalbecas = asignados.count()
                totalrenovacion = asignados.filter(tipo=2).count()
                asignados = asignados.exclude(tipo=2)

                nocargadocumento = asignados.filter(cargadocumento=False).count()
                totaldoccargada = asignados.filter(cargadocumento=True,
                                                   solicitud__inscripcion__persona__personadocumentopersonal__status=True).count()
                totaldocpendcarga = (totalbecas - totalrenovacion - nocargadocumento) - totaldoccargada
                totalnocargadoc = nocargadocumento + totalrenovacion

                documentacion = []
                documentacion.append(['TOTAL ESTUDIANTES QUE NO DEBEN CARGAR DOCUMENTACIÓN', totalnocargadoc])
                documentacion.append(['TOTAL DOCUMENTACIÓN PENDIENTE DE CARGAR POR ESTUDIANTES', totaldocpendcarga])
                documentacion.append(['<strong>TOTAL DOCUMENTACIÓN CARGADA POR ESTUDIANTES</strong>',
                                      '<strong>' + str(totaldoccargada) + '</strong>'])
                data['documentacion'] = documentacion

                validaciondocumento = []
                totaldocvalida1 = asignados.filter(cargadocumento=True,
                                                   solicitud__inscripcion__persona__personadocumentopersonal__status=True,
                                                   solicitud__inscripcion__persona__personadocumentopersonal__estadocedula=2,
                                                   solicitud__inscripcion__persona__personadocumentopersonal__estadopapeleta=2,
                                                   solicitud__inscripcion__persona__personadocumentopersonal__estadocedularepresentantesol=2,
                                                   solicitud__inscripcion__persona__personadocumentopersonal__estadopapeletarepresentantesol=2).exclude(
                    solicitud__becatipo_id=16).count()

                totaldocvalida2 = asignados.filter(cargadocumento=True,
                                                   solicitud__inscripcion__persona__personadocumentopersonal__status=True,
                                                   solicitud__inscripcion__persona__personadocumentopersonal__estadocedula=2,
                                                   solicitud__inscripcion__persona__personadocumentopersonal__estadopapeleta=2,
                                                   solicitud__inscripcion__persona__personadocumentopersonal__estadocedularepresentantesol=2,
                                                   solicitud__inscripcion__persona__personadocumentopersonal__estadopapeletarepresentantesol=2,
                                                   solicitud__inscripcion__persona__personadocumentopersonal__estadoactagrado=2,
                                                   solicitud__becatipo_id=16).count()
                totaldocvalida = totaldocvalida1 + totaldocvalida2

                totaldocrechazada1 = asignados.filter(
                    Q(solicitud__inscripcion__persona__personadocumentopersonal__estadocedula=3) |
                    Q(solicitud__inscripcion__persona__personadocumentopersonal__estadopapeleta=3) |
                    Q(solicitud__inscripcion__persona__personadocumentopersonal__estadocedularepresentantesol=3) |
                    Q(solicitud__inscripcion__persona__personadocumentopersonal__estadopapeletarepresentantesol=3),
                    cargadocumento=True,
                    solicitud__inscripcion__persona__personadocumentopersonal__status=True).exclude(
                    solicitud__becatipo_id=16).count()

                totaldocrechazada2 = asignados.filter(
                    Q(solicitud__inscripcion__persona__personadocumentopersonal__estadocedula=3) |
                    Q(solicitud__inscripcion__persona__personadocumentopersonal__estadopapeleta=3) |
                    Q(solicitud__inscripcion__persona__personadocumentopersonal__estadocedularepresentantesol=3) |
                    Q(solicitud__inscripcion__persona__personadocumentopersonal__estadopapeletarepresentantesol=3) |
                    Q(solicitud__inscripcion__persona__personadocumentopersonal__estadoactagrado=3),
                    cargadocumento=True,
                    solicitud__inscripcion__persona__personadocumentopersonal__status=True,
                    solicitud__becatipo_id=16).count()

                totaldocrechazada = totaldocrechazada1 + totaldocrechazada2
                totaldocpendrevision = totaldoccargada - (totaldocvalida + totaldocrechazada)

                validaciondocumento.append(['TOTAL DOCUMENTACIÓN VALIDADA POR BIENESTAR', totaldocvalida])
                validaciondocumento.append(['TOTAL DOCUMENTACIÓN RECHAZADA POR BIENESTAR', totaldocrechazada])
                validaciondocumento.append(
                    ['TOTAL DOCUMENTACIÓN PENDIENTE DE REVISAR POR BIENESTAR', totaldocpendrevision])
                data['validaciondocumento'] = validaciondocumento
                data['totaldocumentos'] = totaldoccargada

                contratos = []
                # totalcontgen = asignados.filter(numerocontrato__isnull=False, cargadocumento=True).count()
                totalcontgen = asignados.filter(numerocontrato__isnull=False).count()
                totalcontratospendgen = (totaldocvalida + nocargadocumento) - totalcontgen

                contratos1 = asignados.filter(~Q(archivocontrato=''),
                                              archivocontrato__isnull=False,
                                              solicitud__inscripcion__persona__personadocumentopersonal__status=True,
                                              solicitud__inscripcion__persona__personadocumentopersonal__estadocedula=2,
                                              solicitud__inscripcion__persona__personadocumentopersonal__estadopapeleta=2,
                                              solicitud__inscripcion__persona__personadocumentopersonal__estadocedularepresentantesol=2,
                                              solicitud__inscripcion__persona__personadocumentopersonal__estadopapeletarepresentantesol=2,
                                              numerocontrato__isnull=False, cargadocumento=True, tipo=1).exclude(
                    solicitud__becatipo_id=16)

                contratos2 = asignados.filter(~Q(archivocontrato=''),
                                              archivocontrato__isnull=False,
                                              solicitud__inscripcion__persona__personadocumentopersonal__status=True,
                                              solicitud__inscripcion__persona__personadocumentopersonal__estadocedula=2,
                                              solicitud__inscripcion__persona__personadocumentopersonal__estadopapeleta=2,
                                              solicitud__inscripcion__persona__personadocumentopersonal__estadocedularepresentantesol=2,
                                              solicitud__inscripcion__persona__personadocumentopersonal__estadopapeletarepresentantesol=2,
                                              solicitud__inscripcion__persona__personadocumentopersonal__estadoactagrado=2,
                                              numerocontrato__isnull=False, cargadocumento=True,
                                              solicitud__becatipo_id=16, tipo=1)

                contratos3 = asignados.filter(~Q(archivocontrato=''),
                                              archivocontrato__isnull=False,
                                              solicitud__inscripcion__persona__personadocumentopersonal__status=True,
                                              numerocontrato__isnull=False, cargadocumento=False, tipo=1)

                contratoscargados = contratos1 | contratos2 | contratos3

                totalcontcargado = contratoscargados.count()
                totalcontratospencarga = totalcontgen - totalcontcargado

                totalcontratosvalida = contratoscargados.filter(estadorevisioncontrato=2).count()
                totalcontratosrechazado = contratoscargados.filter(estadorevisioncontrato=3).count()
                totalcontratospenrevision = contratoscargados.filter(estadorevisioncontrato=1).count()

                contratos.append(['TOTAL CONTRATOS GENERADOS POR ESTUDIANTES', totalcontgen])
                contratos.append(['TOTAL CONTRATOS PENDIENTES DE GENERAR POR ESTUDIANTES', totalcontratospendgen])
                contratos.append(['<strong>TOTAL CONTRATOS FIRMADOS CARGADOS</strong>',
                                  '<strong>' + str(totalcontcargado) + '</strong>'])
                contratos.append(['TOTAL CONTRATOS FIRMADOS PENDIENTES DE CARGAR', totalcontratospencarga])
                data['contratos'] = contratos

                validacioncontrato = []
                validacioncontrato.append(['TOTAL CONTRATOS VALIDADOS POR BIENESTAR', totalcontratosvalida])
                validacioncontrato.append(['TOTAL CONTRATOS RECHAZADOS POR BIENESTAR', totalcontratosrechazado])
                validacioncontrato.append(
                    ['TOTAL CONTRATOS PENDIENTE DE REVISAR POR BIENESTAR', totalcontratospenrevision])
                data['validacioncontrato'] = validacioncontrato
                data['totalcontratos'] = totalcontratosvalida + totalcontratosrechazado + totalcontratospenrevision

                cuentasbancarias = []
                beneficiarios = Persona.objects.values('id').filter(cuentabancariapersona__status=True,
                                                                    cuentabancariapersona__archivo__isnull=False,
                                                                    inscripcion__becasolicitud__periodo=periodo,
                                                                    inscripcion__becasolicitud__becaasignacion__status=True,
                                                                    inscripcion__becasolicitud__becaasignacion__cargadocumento=True)
                total = beneficiarios.count()
                aprobadas = beneficiarios.filter(cuentabancariapersona__estadorevision=2).count()
                rechazadas = beneficiarios.filter(cuentabancariapersona__estadorevision=3).count()
                totalpendiente = total - (aprobadas + rechazadas)

                cuentasbancarias.append(['TOTAL CUENTAS BANCARIAS REGISTRADAS POR ESTUDIANTES', total])
                cuentasbancarias.append(['TOTAL CUENTAS BANCARIAS VALIDADAS POR TESORERIA', aprobadas])
                cuentasbancarias.append(['TOTAL CUENTAS BANCARIAS RECHAZADAS POR TESORERIA', rechazadas])
                cuentasbancarias.append(['TOTAL CUENTAS BANCARIAS PENDIENTES DE REVISAR POR TESORERIA', totalpendiente])
                data['cuentasbancarias'] = cuentasbancarias

                return conviert_html_to_pdf(
                    'adm_becas/seguimientocontrolbecapdf.html',
                    {
                        'pagesize': 'A4',
                        'data': data,
                    }
                )
            except Exception as ex:
                return HttpResponseRedirect(
                    "/adm_becas?info=%s" % "Error al generar el reporte de seguimiento de asignación de becas")

        elif action == 'reporteseguimientocontrolmayosep2020':
            try:
                data = {}

                becados = BecaAsignacion.objects.filter(status=True,
                                                        solicitud__periodo=periodo,
                                                        solicitudpagobecadetalle__status=True,
                                                        solicitudpagobecadetalle__acreditado=True,
                                                        becacomprobanterevision__isnull=False,
                                                        becacomprobanterevision__status=True
                                                        ).order_by('solicitud__inscripcion__persona__apellido1',
                                                                   'solicitud__inscripcion__persona__apellido2',
                                                                   'solicitud__inscripcion__persona__nombres')

                listado = []
                total_cumple_comp = total_nocumple_comp = total_cumple_apro = total_nocumple_apro = 0
                for beca in becados:
                    alumno = beca.solicitud.inscripcion.persona
                    montoasignado = beca.montobeneficio
                    montofactura = beca.total_monto_comprobantes_cargados()
                    if montofactura >= montoasignado:
                        porc_compra = 100.00
                    else:
                        porc_compra = null_to_decimal((montofactura * 100) / montoasignado, 2)

                    if porc_compra >= 70:
                        cumple_porc_compra = "SI"
                        total_cumple_comp += 1
                    else:
                        cumple_porc_compra = "NO"
                        total_nocumple_comp += 1

                    matricula = beca.solicitud.inscripcion.matricula_periodo_actual(periodo)[0]
                    estado_matricula = "MATRICULADO" if not matricula.retirado() else "RETIRADO"
                    porc_materia_apro = matricula.porcentaje_materias_aprobadas() if not matricula.retirado() else 0
                    if porc_materia_apro >= 70:
                        cumple_porc_apro = "SI"
                        total_cumple_apro += 1
                    else:
                        cumple_porc_apro = "NO"
                        total_nocumple_apro += 1

                    listado.append(
                        [alumno.identificacion(), alumno.nombre_completo_inverso(), estado_matricula, porc_compra,
                         cumple_porc_compra, porc_materia_apro, cumple_porc_apro])

                data['periodo'] = periodo.nombre
                data['listado'] = listado
                data['total_cumple_comp'] = total_cumple_comp
                data['total_nocumple_comp'] = total_nocumple_comp
                data['total_cumple_apro'] = total_cumple_apro
                data['total_nocumple_apro'] = total_nocumple_apro
                return conviert_html_to_pdf(
                    'adm_becas/seguimientocontrolms2020becapdf.html',
                    {
                        'pagesize': 'A4',
                        'data': data,
                    }
                )
            except Exception as ex:
                return HttpResponseRedirect(
                    "/adm_becas?info=%s" % "Error al generar el reporte de seguimiento y control")

        elif action == 'listadoseguimientocontrolmayosep2020':
            try:
                __author__ = 'Unemi'

                titulo = easyxf(
                    'font: name Times New Roman, color-index black, bold on , height 350; alignment: horiz centre')
                titulo2 = easyxf(
                    'font: name Times New Roman, color-index black, bold on , height 250; alignment: horiz centre')
                fuentecabecera = easyxf(
                    'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
                fuentenormal = easyxf(
                    'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
                fuentenumeroentero = easyxf(
                    'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right')
                fuenteporcentaje = easyxf(
                    'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right',
                    num_format_str='##0.00%')

                font_style = XFStyle()
                font_style.font.bold = True
                font_style2 = XFStyle()
                font_style2.font.bold = False
                wb = Workbook(encoding='utf-8')
                ws = wb.add_sheet('Becados')
                response = HttpResponse(content_type="application/ms-excel")
                response['Content-Disposition'] = 'attachment; filename=becados_seguimiento_control_' + random.randint(
                    1,
                    10000).__str__() + '.xls'

                ws.write_merge(0, 0, 0, 7, 'UNIVERSIDAD ESTATAL DE MILAGRO', titulo)
                ws.write_merge(1, 1, 0, 7, 'DIRECCIÓN DE BIENESTAR UNIVERSITARIO', titulo2)
                ws.write_merge(2, 2, 0, 7, 'LISTADO DE BECADOS - SEGUIMIENTO Y CONTROL DE AYUDAS ECONÓMICAS', titulo2)
                ws.write_merge(3, 3, 0, 7, 'PERIODO ' + periodo.nombre, titulo2)

                ws.write_merge(5, 6, 0, 0, u'#', fuentecabecera)
                ws.write_merge(5, 6, 1, 1, u'IDENTIFICACIÓN', fuentecabecera)
                ws.write_merge(5, 6, 2, 2, u'APELLIDOS/NOMBRES', fuentecabecera)
                ws.write_merge(5, 6, 3, 3, u'ESTADO MATRICULA', fuentecabecera)
                ws.write_merge(5, 5, 4, 7, u'CUMPLIMIENTO', fuentecabecera)
                ws.write(6, 4, "70 % VALOR COMPRA", fuentecabecera)
                ws.write(6, 5, "CUMPLE", fuentecabecera)
                ws.write(6, 6, "70 % MATERIAS APROBADAS", fuentecabecera)
                ws.write(6, 7, "CUMPLE", fuentecabecera)

                row_num = 6

                becados = BecaAsignacion.objects.filter(status=True,
                                                        solicitud__periodo=periodo,
                                                        solicitudpagobecadetalle__status=True,
                                                        solicitudpagobecadetalle__acreditado=True,
                                                        becacomprobanterevision__isnull=False,
                                                        becacomprobanterevision__status=True
                                                        ).order_by('solicitud__inscripcion__persona__apellido1',
                                                                   'solicitud__inscripcion__persona__apellido2',
                                                                   'solicitud__inscripcion__persona__nombres')
                c = 0
                for beca in becados:
                    row_num += 1
                    c += 1
                    alumno = beca.solicitud.inscripcion.persona
                    montoasignado = beca.montobeneficio
                    montofactura = beca.total_monto_comprobantes_cargados()
                    if montofactura >= montoasignado:
                        porc_compra = 100.00
                    else:
                        porc_compra = null_to_decimal((montofactura * 100) / montoasignado, 2)

                    if porc_compra >= 70:
                        cumple_porc_compra = "SI"
                    else:
                        cumple_porc_compra = "NO"

                    porc_compra /= 100

                    matricula = beca.solicitud.inscripcion.matricula_periodo_actual(periodo)[0]
                    estado_matricula = "MATRICULADO" if not matricula.retirado() else "RETIRADO"
                    porc_materia_apro = matricula.porcentaje_materias_aprobadas() if not matricula.retirado() else 0
                    if porc_materia_apro >= 70:
                        cumple_porc_apro = "SI"
                    else:
                        cumple_porc_apro = "NO"

                    porc_materia_apro /= 100

                    ws.write(row_num, 0, c, fuentenumeroentero)
                    ws.write(row_num, 1, alumno.identificacion(), fuentenormal)
                    ws.write(row_num, 2, alumno.nombre_completo_inverso(), fuentenormal)
                    ws.write(row_num, 3, estado_matricula, fuentenormal)
                    ws.write(row_num, 4, porc_compra, fuenteporcentaje)
                    ws.write(row_num, 5, cumple_porc_compra, fuentenormal)
                    ws.write(row_num, 6, porc_materia_apro, fuenteporcentaje)
                    ws.write(row_num, 7, cumple_porc_apro, fuentenormal)
                    print(c)

                row_num += 2
                for col_num in range(8):
                    if col_num == 0:
                        ws.col(col_num).width = 1500
                    elif col_num == 1:
                        ws.col(col_num).width = 4000
                    elif col_num == 2:
                        ws.col(col_num).width = 10000
                    elif col_num == 3:
                        ws.col(col_num).width = 3000
                    elif col_num == 4 or col_num == 6:
                        ws.col(col_num).width = 4000
                    else:
                        ws.col(col_num).width = 2000

                wb.save(response)
                return response
            except Exception as ex:
                print("Error...")
                pass

        elif action == 'listadoseguimientocontrolnov2019mar2020':
            try:
                __author__ = 'Unemi'

                titulo = easyxf(
                    'font: name Times New Roman, color-index black, bold on , height 350; alignment: horiz centre')
                titulo2 = easyxf(
                    'font: name Times New Roman, color-index black, bold on , height 250; alignment: horiz centre')
                fuentecabecera = easyxf(
                    'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
                fuentenormal = easyxf(
                    'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
                fuentenumeroentero = easyxf(
                    'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right')
                fuenteporcentaje = easyxf(
                    'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right',
                    num_format_str='##0.00%')

                font_style = XFStyle()
                font_style.font.bold = True
                font_style2 = XFStyle()
                font_style2.font.bold = False
                wb = Workbook(encoding='utf-8')
                ws = wb.add_sheet('Becados')
                response = HttpResponse(content_type="application/ms-excel")
                response['Content-Disposition'] = 'attachment; filename=becados_seguimiento_control_' + random.randint(
                    1, 10000).__str__() + '.xls'

                ws.write_merge(0, 0, 0, 7, 'UNIVERSIDAD ESTATAL DE MILAGRO', titulo)
                ws.write_merge(1, 1, 0, 7, 'DIRECCIÓN DE BIENESTAR UNIVERSITARIO', titulo2)
                ws.write_merge(2, 2, 0, 7, 'LISTADO DE BECADOS - SEGUIMIENTO Y CONTROL DE BECAS', titulo2)
                ws.write_merge(3, 3, 0, 7, 'PERIODO ' + periodo.nombre, titulo2)

                row_num = 5

                ws.write_merge(5, 6, 0, 0, u'#', fuentecabecera)
                ws.write_merge(5, 6, 1, 1, u'IDENTIFICACIÓN', fuentecabecera)
                ws.write_merge(5, 6, 2, 2, u'APELLIDOS/NOMBRES', fuentecabecera)
                ws.write_merge(5, 6, 3, 3, u'ESTADO MATRICULA', fuentecabecera)
                ws.write_merge(5, 5, 4, 7, u'% ASISTENCIA', fuentecabecera)
                ws.write(6, 4, "NOVIEMBRE", fuentecabecera)
                ws.write(6, 5, "DICIEMBRE", fuentecabecera)
                ws.write(6, 6, "ENERO", fuentecabecera)
                ws.write(6, 7, "FEBRERO", fuentecabecera)

                row_num = 6

                becados = BecaAsignacion.objects.filter(status=True,
                                                        solicitud__periodo=periodo,
                                                        solicitudpagobecadetalle__asignacion__isnull=False
                                                        ).order_by('solicitud__inscripcion__persona__apellido1',
                                                                   'solicitud__inscripcion__persona__apellido2',
                                                                   'solicitud__inscripcion__persona__nombres')
                c = 0
                for beca in becados:
                    row_num += 1
                    c += 1
                    alumno = beca.solicitud.inscripcion.persona

                    matricula = beca.solicitud.inscripcion.matricula_periodo_actual(periodo)[0]
                    estado_matricula = "MATRICULADO" if not matricula.retirado() else "RETIRADO"

                    # Asistencia mes 1
                    desde = datetime.strptime('2019-11-01', '%Y-%m-%d').date()
                    hasta = datetime.strptime('2019-11-30', '%Y-%m-%d').date()
                    porcmes1 = beca.porcentaje_asistencia_mes(desde, hasta)

                    # Asistencia mes 2
                    desde = datetime.strptime('2019-12-01', '%Y-%m-%d').date()
                    hasta = datetime.strptime('2019-12-31', '%Y-%m-%d').date()
                    porcmes2 = beca.porcentaje_asistencia_mes(desde, hasta)

                    # Asistencia mes 3
                    desde = datetime.strptime('2020-01-01', '%Y-%m-%d').date()
                    hasta = datetime.strptime('2020-01-31', '%Y-%m-%d').date()
                    porcmes3 = beca.porcentaje_asistencia_mes(desde, hasta)

                    # Asistencia mes 4
                    desde = datetime.strptime('2020-02-01', '%Y-%m-%d').date()
                    hasta = datetime.strptime('2020-02-29', '%Y-%m-%d').date()
                    porcmes4 = beca.porcentaje_asistencia_mes(desde, hasta)

                    ws.write(row_num, 0, c, fuentenumeroentero)
                    ws.write(row_num, 1, alumno.identificacion(), fuentenormal)
                    ws.write(row_num, 2, alumno.nombre_completo_inverso(), fuentenormal)
                    ws.write(row_num, 3, estado_matricula, fuentenormal)
                    ws.write(row_num, 4, porcmes1 / 100, fuenteporcentaje)
                    ws.write(row_num, 5, porcmes2 / 100, fuenteporcentaje)
                    ws.write(row_num, 6, porcmes3 / 100, fuenteporcentaje)
                    ws.write(row_num, 7, porcmes4 / 100, fuenteporcentaje)
                    print(c)

                row_num += 2
                for col_num in range(7):
                    if col_num == 0:
                        ws.col(col_num).width = 1500
                    elif col_num == 1:
                        ws.col(col_num).width = 4000
                    elif col_num == 2:
                        ws.col(col_num).width = 10000
                    else:
                        ws.col(col_num).width = 3000

                wb.save(response)
                return response
            except Exception as ex:
                print("Error...")
                pass

        elif action == 'reporteseguimientocontrolnov2019mar2020':
            try:
                data = {}

                becados = BecaAsignacion.objects.filter(status=True, solicitud__periodo=periodo,
                                                        solicitudpagobecadetalle__asignacion__isnull=False).order_by(
                    'solicitud__inscripcion__persona__apellido1',
                    'solicitud__inscripcion__persona__apellido2',
                    'solicitud__inscripcion__persona__nombres')

                listado = []

                for beca in becados:
                    alumno = beca.solicitud.inscripcion.persona

                    matricula = beca.solicitud.inscripcion.matricula_periodo_actual(periodo)[0]
                    estado_matricula = "MATRICULADO" if not matricula.retirado() else "RETIRADO"

                    # Asistencia mes 1
                    desde = datetime.strptime('2019-11-01', '%Y-%m-%d').date()
                    hasta = datetime.strptime('2019-11-30', '%Y-%m-%d').date()
                    porcmes1 = beca.porcentaje_asistencia_mes(desde, hasta)

                    # Asistencia mes 2
                    desde = datetime.strptime('2019-12-01', '%Y-%m-%d').date()
                    hasta = datetime.strptime('2019-12-31', '%Y-%m-%d').date()
                    porcmes2 = beca.porcentaje_asistencia_mes(desde, hasta)

                    # Asistencia mes 3
                    desde = datetime.strptime('2020-01-01', '%Y-%m-%d').date()
                    hasta = datetime.strptime('2020-01-31', '%Y-%m-%d').date()
                    porcmes3 = beca.porcentaje_asistencia_mes(desde, hasta)

                    # Asistencia mes 4
                    desde = datetime.strptime('2020-02-01', '%Y-%m-%d').date()
                    hasta = datetime.strptime('2020-02-29', '%Y-%m-%d').date()
                    porcmes4 = beca.porcentaje_asistencia_mes(desde, hasta)

                    listado.append(
                        [alumno.identificacion(), alumno.nombre_completo_inverso(), estado_matricula, porcmes1,
                         porcmes2, porcmes3, porcmes4])

                data['periodo'] = periodo.nombre
                data['listado'] = listado

                return conviert_html_to_pdf(
                    'adm_becas/seguimientocontrolo2019m2020becapdf.html',
                    {
                        'pagesize': 'A4',
                        'data': data,
                    }
                )
            except Exception as ex:
                return HttpResponseRedirect(
                    "/adm_becas?info=%s" % "Error al generar el reporte de seguimiento y control")

        elif action == 'listadosolicitudpago':
            try:
                __author__ = 'Unemi'
                solicitudpago = SolicitudPagoBeca.objects.get(pk=int(request.POST['id']))
                numeroreporte = "Nº " + str(solicitudpago.id).zfill(5)

                titulo = easyxf(
                    'font: name Times New Roman, color-index black, bold on , height 350; alignment: horiz centre')
                titulo2 = easyxf(
                    'font: name Times New Roman, color-index black, bold on , height 250; alignment: horiz centre')
                fuentecabecera = easyxf(
                    'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
                fuentenormal = easyxf(
                    'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
                fuentemoneda = easyxf(
                    'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right',
                    num_format_str=' "$" #,##0.00')
                fuentefecha = easyxf(
                    'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz center',
                    num_format_str='yyyy-mm-dd')
                fuentenumerodecimal = easyxf(
                    'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right',
                    num_format_str='#,##0.00')
                fuentenumeroentero = easyxf(
                    'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right')

                font_style = XFStyle()
                font_style.font.bold = True
                font_style2 = XFStyle()
                font_style2.font.bold = False
                wb = Workbook(encoding='utf-8')
                ws = wb.add_sheet('Becados')
                response = HttpResponse(content_type="application/ms-excel")
                response['Content-Disposition'] = 'attachment; filename=becados_solicitudpago_' + random.randint(1,
                                                                                                                 10000).__str__() + '.xls'

                ws.write_merge(0, 0, 0, 12, 'UNIVERSIDAD ESTATAL DE MILAGRO', titulo)
                ws.write_merge(1, 1, 0, 12, 'DIRECCIÓN DE BIENESTAR UNIVERSITARIO', titulo2)
                ws.write_merge(2, 2, 0, 12, 'LISTADO DE BECADOS - SOLICITUD DE PAGO - ' + numeroreporte, titulo2)

                row_num = 4
                columns = [
                    (u"# SOLICITUD", 3000),
                    (u"FECHA BECA", 3000),
                    # (u"MONTO BECA", 3000),
                    (u"NOMBRES COMPLETOS", 10000),
                    (u"IDENTIFICACIÓN", 5000),
                    (u"FACULTAD", 5000),
                    (u"CARRERA", 5000),
                    (u"MODALIDAD", 5000),
                    (u"NIVEL", 5000),
                    (u"PROVINCIA", 5000),
                    (u"CANTÓN", 5000),
                    (u"PARROQUIA", 5000),
                    (u"DIRECCIÓN", 15000),
                    (u"REFERENCIA", 15000),
                    (u"SECTOR", 15000),
                    (u"# CASA", 5000),
                    (u"CORREO PERSONAL", 8000),
                    (u"CORREO UNEMI", 8000),
                    (u"TELÉFONO", 5000),
                    (u"CELULAR", 5000),
                    (u"OPERADORA", 5000)
                ]
                for col_num in range(len(columns)):
                    ws.write(row_num, col_num, columns[col_num][0], fuentecabecera)
                    ws.col(col_num).width = columns[col_num][1]

                row_num = 4

                detallesolicitud = solicitudpago.solicitudpagobecadetalle_set.all().order_by(
                    'asignacion__solicitud__inscripcion__persona__apellido1',
                    'asignacion__solicitud__inscripcion__persona__apellido2',
                    'asignacion__solicitud__inscripcion__persona__nombres')
                # detallesolicitud = detallesolicitud[:500]
                # data['fechasolicitud'] = str(solicitudpago.fecha.day) + " de " + \
                #                          MESES_CHOICES[solicitudpago.fecha.month - 1][
                #                              1].capitalize() + " del " + str(solicitudpago.fecha.year)

                for s in detallesolicitud:
                    row_num += 1
                    p = Persona.objects.get(pk=s.asignacion.solicitud.inscripcion.persona_id)

                    umat = None
                    if s.asignacion.solicitud.inscripcion.matricula_periodo_actual(periodo):
                        umat = s.asignacion.solicitud.inscripcion.matricula_periodo_actual(periodo)[0]

                    ws.write(row_num, 0, str(s.asignacion.solicitud.id), fuentenumeroentero)
                    ws.write(row_num, 1, str(s.asignacion.fecha_creacion)[:10], fuentefecha)
                    # ws.write(row_num, 2, s.asignacion.montobeneficio, fuentemoneda)
                    ws.write(row_num, 2, p.nombre_completo_inverso(), fuentenormal)
                    ws.write(row_num, 3, p.identificacion(), fuentenormal)
                    ws.write(row_num, 4, str(umat.nivel.coordinacion()) if umat else '', fuentenormal)
                    ws.write(row_num, 5, str(umat.inscripcion.carrera) if umat else '', fuentenormal)
                    ws.write(row_num, 6, str(umat.inscripcion.modalidad) if umat else '', fuentenormal)
                    ws.write(row_num, 7, umat.nivelmalla.nombre if umat else '', fuentenormal)
                    ws.write(row_num, 8, str(p.provincia) if p.provincia else '', fuentenormal)
                    ws.write(row_num, 9, str(p.canton) if p.canton else '', fuentenormal)
                    ws.write(row_num, 10, str(p.parroquia) if p.parroquia else '', fuentenormal)
                    ws.write(row_num, 11, p.direccion_corta().upper(), fuentenormal)
                    ws.write(row_num, 12, p.referencia.upper(), fuentenormal)
                    ws.write(row_num, 13, p.sector.upper(), fuentenormal)
                    ws.write(row_num, 14, p.num_direccion, fuentenormal)
                    ws.write(row_num, 15, p.email, fuentenormal)
                    ws.write(row_num, 16, p.emailinst, fuentenormal)
                    ws.write(row_num, 17, p.telefono_conv, fuentenormal)
                    ws.write(row_num, 18, p.telefono, fuentenormal)
                    ws.write(row_num, 19, p.get_tipocelular_display() if p.tipocelular else '', fuentenormal)

                wb.save(response)
                return response
            except Exception as ex:
                print("Error...")
                pass

        elif action == 'verificar_solicitudes':
            try:
                if datetime.strptime(request.POST['desde'], '%d-%m-%Y') <= datetime.strptime(request.POST['hasta'],
                                                                                             '%d-%m-%Y'):
                    desde = datetime.combine(convertir_fecha(request.POST['desde']), time.min)
                    hasta = datetime.combine(convertir_fecha(request.POST['hasta']), time.max)
                    tipobeca = int(request.POST['tipobeca'])

                    if BecaSolicitud.objects.filter(fecha_creacion__range=(desde, hasta), status=True, periodo=periodo,
                                                    becatipo_id=tipobeca).exclude(estado__in=[5, 8]).exists():
                        return JsonResponse({"result": "ok"})
                    else:
                        return JsonResponse(
                            {"result": "bad", "mensaje": "No existen registros para generar el reporte"})
                else:
                    return JsonResponse(
                        {"result": "bad", "mensaje": "La fecha desde debe ser menor o igual a la fecha hasta"})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al generar la consulta."})

        elif action == 'verificar_solicitudes_desistidas':
            try:
                if BecaSolicitud.objects.filter(status=True, periodo=periodo, estado=8).exists():
                    return JsonResponse({"result": "ok"})
                else:
                    return JsonResponse({"result": "bad", "mensaje": "No existen registros para generar el reporte"})

            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al generar la consulta."})

        elif action == 'verificar_solicitudes_pendientes_aceptar':
            try:
                if BecaSolicitud.objects.filter(status=True, periodo=periodo, estado=2, becaaceptada=1).exists():
                    return JsonResponse({"result": "ok"})
                else:
                    return JsonResponse({"result": "bad", "mensaje": "No existen registros para generar el reporte"})

            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al generar la consulta."})

        elif action == 'verificar_solicitudes_pendiente_revisar':
            try:
                if BecaSolicitud.objects.filter(status=True, periodo=periodo, estado__in=[1, 4],
                                                periodocalifica__isnull=False).exists():
                    return JsonResponse({"result": "ok"})
                else:
                    return JsonResponse({"result": "bad", "mensaje": "No existen registros para generar el reporte"})

            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al generar la consulta."})

        elif action == 'verificar_listado_beneficiarios':
            try:
                if BecaSolicitud.objects.filter(status=True, periodo=periodo).exists():
                    return JsonResponse({"result": "ok"})
                else:
                    return JsonResponse({"result": "bad", "mensaje": "No existen registros para generar el reporte"})

            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al generar la consulta."})

        elif action == 'verificar_becas_asignadas':
            try:
                from datetime import time
                if datetime.strptime(request.POST['desde'], '%d-%m-%Y') <= datetime.strptime(request.POST['hasta'],
                                                                                             '%d-%m-%Y'):
                    desde = datetime.combine(convertir_fecha(request.POST['desde']), time.min)
                    hasta = datetime.combine(convertir_fecha(request.POST['hasta']), time.max)
                    tipobeca = int(request.POST['tipobeca'])

                    # if tipobeca == 0:
                    #     tipos = [18, 19]
                    # else:
                    #     tipos = [tipobeca]

                    # if BecaAsignacion.objects.filter(fecha_creacion__range=(desde, hasta), status=True, solicitud__periodo=periodo, solicitud__becatipo_id__in=tipos).exists():
                    becas1 = BecaAsignacion.objects.filter(cargadocumento=True,
                                                           solicitud__inscripcion__persona__personadocumentopersonal__estadocedula=2,
                                                           solicitud__inscripcion__persona__personadocumentopersonal__estadocedularepresentantesol=2,
                                                           solicitud__inscripcion__persona__personadocumentopersonal__estadopapeleta=2,
                                                           solicitud__inscripcion__persona__personadocumentopersonal__estadopapeletarepresentantesol=2,
                                                           solicitud__inscripcion__persona__personadocumentopersonal__fecha_modificacion__range=(
                                                               desde, hasta),
                                                           status=True, solicitud__periodo=periodo, tipo=1).exclude(
                        solicitud__becatipo_id=16)

                    becas2 = BecaAsignacion.objects.filter(cargadocumento=True,
                                                           solicitud__inscripcion__persona__personadocumentopersonal__estadocedula=2,
                                                           solicitud__inscripcion__persona__personadocumentopersonal__estadocedularepresentantesol=2,
                                                           solicitud__inscripcion__persona__personadocumentopersonal__estadopapeleta=2,
                                                           solicitud__inscripcion__persona__personadocumentopersonal__estadopapeletarepresentantesol=2,
                                                           solicitud__inscripcion__persona__personadocumentopersonal__estadoactagrado=2,
                                                           solicitud__inscripcion__persona__personadocumentopersonal__fecha_modificacion__range=(
                                                               desde, hasta),
                                                           status=True, solicitud__periodo=periodo,
                                                           solicitud__becatipo_id=16, tipo=1)

                    if periodo.id >= 119:
                        becas1 = BecaAsignacion.objects.filter(cargadocumento=True,
                                                               solicitud__inscripcion__persona__personadocumentopersonal__estadocedula=2,
                                                               # solicitud__inscripcion__persona__personadocumentopersonal__estadocedularepresentantesol=2,
                                                               # solicitud__inscripcion__persona__personadocumentopersonal__estadopapeleta=2,
                                                               # solicitud__inscripcion__persona__personadocumentopersonal__estadopapeletarepresentantesol=2,
                                                               solicitud__inscripcion__persona__personadocumentopersonal__fecha_modificacion__range=(
                                                                   desde, hasta),
                                                               status=True, solicitud__periodo=periodo, tipo=1).exclude(solicitud__becatipo_id=16)

                        becas2 = BecaAsignacion.objects.filter(cargadocumento=True,
                                                               solicitud__inscripcion__persona__personadocumentopersonal__estadocedula=2,
                                                               # solicitud__inscripcion__persona__personadocumentopersonal__estadocedularepresentantesol=2,
                                                               # solicitud__inscripcion__persona__personadocumentopersonal__estadopapeleta=2,
                                                               # solicitud__inscripcion__persona__personadocumentopersonal__estadopapeletarepresentantesol=2,
                                                               solicitud__inscripcion__persona__personadocumentopersonal__estadoactagrado=2,
                                                               solicitud__inscripcion__persona__personadocumentopersonal__fecha_modificacion__range=(desde, hasta),
                                                               status=True, solicitud__periodo=periodo,
                                                               solicitud__becatipo_id=16, tipo=1)

                    becas = becas1 | becas2

                    if becas:
                        return JsonResponse({"result": "ok"})
                    else:
                        return JsonResponse({"result": "bad",
                                             "mensaje": "No existen registros de becas con documentación validada para generar el reporte"})
                else:
                    return JsonResponse(
                        {"result": "bad", "mensaje": "La fecha desde debe ser menor o igual a la fecha hasta"})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al generar la consulta."})

        elif action == 'verificar_becas_documentacion_pendiente':
            try:
                if BecaAsignacion.objects.filter((Q(solicitud__inscripcion__persona__personadocumentopersonal__isnull=True) |
                                                  Q(solicitud__inscripcion__persona__cuentabancariapersona__isnull=True)),
                                                 status=True, solicitud__periodo=periodo,
                                                 tipo=1, cargadocumento=True
                                                 ).exists():
                    return JsonResponse({"result": "ok"})
                else:
                    return JsonResponse({"result": "bad",
                                         "mensaje": "No existen registros de becas con documentación pendiente de cargar para generar el reporte"})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al generar la consulta."})

        elif action == 'verificar_becas_matriz_siies_caces':
            try:
                nueva = BecaAsignacion.objects.values('id').filter(status=True, solicitud__periodo=periodo,
                                                                   solicitud__becatipo_id__isnull=False,
                                                                   estadorevisioncontrato=2, tipo=1)
                renovacion = BecaAsignacion.objects.values('id').filter(status=True, solicitud__periodo=periodo,
                                                                        solicitud__becatipo_id__isnull=False, tipo=2)

                if nueva or renovacion:
                    return JsonResponse({"result": True})
                else:
                    return JsonResponse({"result": False, "mensaje": "No existen registros de becas"})
            except Exception as ex:
                return JsonResponse({"result": False, "mensaje": u"Error al generar la consulta."})

        elif action == 'verificar_becas_por_tipo':
            try:
                becas = BecaAsignacion.objects.values('id').filter(status=True,
                                                                   solicitud__periodo=periodo,
                                                                   solicitud__becatipo_id__isnull=False,
                                                                   solicitudpagobecadetalle__asignacion__isnull=False)

                if becas:
                    return JsonResponse({"result": "ok"})
                else:
                    return JsonResponse({"result": "bad", "mensaje": "No existen registros de becas"})
            except Exception as ex:
                return JsonResponse({"rreporteseguimientopdfesult": "bad", "mensaje": u"Error al generar la consulta."})

        elif action == 'reportebecaportipopdf':
            try:
                data = {}

                data['periodo'] = periodo.nombre
                fecha = datetime.now().date()
                data['fecha'] = str(fecha.day) + " de " + MESES_CHOICES[fecha.month - 1][
                    1].capitalize() + " del " + str(fecha.year)

                datos = []
                totalbecas = totalmasc = totalfem = 0

                tiposbecas = BecaTipo.objects.values('id', 'nombre').filter(becasolicitud__periodo=periodo,
                                                                            becasolicitud__status=True,
                                                                            becasolicitud__becaasignacion__status=True,
                                                                            becasolicitud__becaasignacion__solicitudpagobecadetalle__asignacion__isnull=False).order_by(
                    'nombre').distinct()
                for tipo in tiposbecas:
                    becas = BecaAsignacion.objects.filter(status=True,
                                                          solicitud__periodo=periodo,
                                                          solicitud__becatipo__id=tipo['id'],
                                                          solicitudpagobecadetalle__asignacion__isnull=False).exclude(
                        estadobeca=2)
                    masc = becas.filter(solicitud__inscripcion__persona__sexo_id=2).count()
                    fem = becas.filter(solicitud__inscripcion__persona__sexo_id=1).count()

                    datos.append([tipo['nombre'].upper(), masc, fem, masc + fem])
                    totalmasc += masc
                    totalfem += fem
                    totalbecas += becas.count()

                data['datos'] = datos
                data['totalbecas'] = totalbecas
                data['totalmasc'] = totalmasc
                data['totalfem'] = totalfem
                return conviert_html_to_pdf(
                    'adm_becas/becasportipopdf.html',
                    {
                        'pagesize': 'A4',
                        'data': data,
                    }
                )
            except Exception as ex:
                return HttpResponseRedirect(
                    "/adm_becas?info=%s" % "Error al generar el reporte de seguimiento de asignación de becas")

        elif action == 'reportebecaporfacultadcarrerapdf':
            try:
                data = {}

                data['periodo'] = periodo.nombre
                fecha = datetime.now().date()
                data['fecha'] = str(fecha.day) + " de " + MESES_CHOICES[fecha.month - 1][
                    1].capitalize() + " del " + str(fecha.year)

                listafacultades = []
                listacarreras = []
                datos = []

                totalbecas = totalmasc = totalfem = 0

                becas = BecaAsignacion.objects.filter(status=True, solicitud__periodo=periodo,
                                                      solicitudpagobecadetalle__asignacion__isnull=False).exclude(
                    estadobeca=2)

                facultades = Coordinacion.objects.values('id', 'nombre', 'alias').filter(
                    inscripcion__becasolicitud__becaasignacion__status=True,
                    inscripcion__becasolicitud__becaasignacion__solicitudpagobecadetalle__asignacion__isnull=False).order_by(
                    'nombre').distinct()

                # Facultades
                for f in facultades:
                    x = becas.filter(solicitud__inscripcion__coordinacion_id=f['id']).count()
                    # thf = Total hombres por facultad, tmf = Total mujeres por facultad
                    thf = becas.filter(solicitud__inscripcion__coordinacion_id=f['id'],
                                       solicitud__inscripcion__persona__sexo_id=2).count()
                    tmf = becas.filter(solicitud__inscripcion__coordinacion_id=f['id'],
                                       solicitud__inscripcion__persona__sexo_id=1).count()
                    listafacultades.append([f['id'], f['nombre'], f['alias'], thf, tmf, thf + tmf, x])

                    # Carreras
                    for c in becas.values('solicitud__inscripcion__carrera_id',
                                          'solicitud__inscripcion__carrera__nombre',
                                          'solicitud__inscripcion__carrera__alias').filter(
                        solicitud__inscripcion__coordinacion_id=f['id']).order_by(
                        'solicitud__inscripcion__carrera__nombre').distinct():
                        # thc = Total hombres carrera, tmc = Total mujeres carrera
                        thc = becas.filter(solicitud__inscripcion__coordinacion_id=f['id'],
                                           solicitud__inscripcion__carrera_id=c['solicitud__inscripcion__carrera_id'],
                                           solicitud__inscripcion__persona__sexo_id=2).count()
                        tmc = becas.filter(solicitud__inscripcion__coordinacion_id=f['id'],
                                           solicitud__inscripcion__carrera_id=c['solicitud__inscripcion__carrera_id'],
                                           solicitud__inscripcion__persona__sexo_id=1).count()
                        listacarreras.append([f['id'], c['solicitud__inscripcion__carrera_id'],
                                              c['solicitud__inscripcion__carrera__nombre'],
                                              c['solicitud__inscripcion__carrera__alias'] if len(
                                                  c['solicitud__inscripcion__carrera__alias']) else '-', thc, tmc,
                                              thc + tmc])

                        tiposbeca = becas.values('solicitud__becatipo_id', 'solicitud__becatipo__nombre').filter(
                            solicitud__inscripcion__coordinacion_id=f['id'],
                            solicitud__inscripcion__carrera_id=c['solicitud__inscripcion__carrera_id']).order_by(
                            'solicitud__becatipo__nombre').distinct()

                        # Tipos de Beca
                        ctb = 1
                        for tipo in tiposbeca:
                            # tht = Total hombres por tipo, tmt = Total mujeres por tipo
                            tht = becas.filter(solicitud__becatipo_id=tipo['solicitud__becatipo_id'],
                                               solicitud__inscripcion__coordinacion_id=f['id'],
                                               solicitud__inscripcion__carrera_id=c[
                                                   'solicitud__inscripcion__carrera_id'],
                                               solicitud__inscripcion__persona__sexo_id=2).count()
                            tmt = becas.filter(solicitud__becatipo_id=tipo['solicitud__becatipo_id'],
                                               solicitud__inscripcion__coordinacion_id=f['id'],
                                               solicitud__inscripcion__carrera_id=c[
                                                   'solicitud__inscripcion__carrera_id'],
                                               solicitud__inscripcion__persona__sexo_id=1).count()

                            datos.append([c['solicitud__inscripcion__carrera_id'], ctb, tipo['solicitud__becatipo_id'],
                                          tipo['solicitud__becatipo__nombre'].upper(), tht, tmt, tht + tmt])
                            ctb += 1

                            totalmasc += tht
                            totalfem += tmt

                data['facultades'] = listafacultades
                data['carreras'] = listacarreras
                data['datos'] = datos
                data['totalbecas'] = totalbecas
                data['totalmasc'] = totalmasc
                data['totalfem'] = totalfem
                return conviert_html_to_pdf(
                    'adm_becas/becasporfacultadcarrerapdf.html',
                    {
                        'pagesize': 'A4',
                        'data': data,
                    }
                )
            except Exception as ex:
                return HttpResponseRedirect(
                    "/adm_becas?info=%s" % "Error al generar el reporte de seguimiento de asignación de becas")

        elif action == 'verificar_becas_contrato_pendiente':
            try:
                becas1 = BecaAsignacion.objects.filter(
                    solicitud__inscripcion__persona__personadocumentopersonal__status=True,
                    solicitud__inscripcion__persona__personadocumentopersonal__estadocedula=2,
                    solicitud__inscripcion__persona__personadocumentopersonal__estadopapeleta=2,
                    solicitud__inscripcion__persona__personadocumentopersonal__estadocedularepresentantesol=2,
                    solicitud__inscripcion__persona__personadocumentopersonal__estadopapeletarepresentantesol=2,
                    numerocontrato__isnull=True,
                    status=True, solicitud__periodo=periodo, tipo=1).exclude(solicitud__becatipo_id=16)

                becas2 = BecaAsignacion.objects.filter(
                    solicitud__inscripcion__persona__personadocumentopersonal__status=True,
                    solicitud__inscripcion__persona__personadocumentopersonal__estadocedula=2,
                    solicitud__inscripcion__persona__personadocumentopersonal__estadopapeleta=2,
                    solicitud__inscripcion__persona__personadocumentopersonal__estadocedularepresentantesol=2,
                    solicitud__inscripcion__persona__personadocumentopersonal__estadopapeletarepresentantesol=2,
                    solicitud__inscripcion__persona__personadocumentopersonal__estadoactagrado=2,
                    numerocontrato__isnull=True,
                    status=True, solicitud__periodo=periodo, tipo=1, solicitud__becatipo_id=16)

                becas = becas1 | becas2

                if becas:
                    return JsonResponse({"result": "ok"})
                else:
                    return JsonResponse({"result": "bad",
                                         "mensaje": "No existen registros de becas con contratos pendientes de generar por los estudiantes para generar el reporte"})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al generar la consulta."})

        elif action == 'verificar_becas_contrato_pendiente_cargar':
            try:
                if BecaAsignacion.objects.filter(Q(archivocontrato='') | Q(archivocontrato__isnull=True),
                                                 numerocontrato__isnull=False,
                                                 status=True, solicitud__periodo=periodo).exists():
                    return JsonResponse({"result": "ok"})
                else:
                    return JsonResponse({"result": "bad",
                                         "mensaje": "No existen registros de becas con contratos pendientes de generar por los estudiantes para generar el reporte"})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al generar la consulta."})

        elif action == 'verificar_becas_contrato_rechazado':
            try:
                if BecaAsignacion.objects.filter(~Q(archivocontrato=''),
                                                 archivocontrato__isnull=False,
                                                 numerocontrato__isnull=False,
                                                 estadorevisioncontrato=3,
                                                 status=True, solicitud__periodo=periodo).exists():
                    return JsonResponse({"result": "ok"})
                else:
                    return JsonResponse({"result": "bad",
                                         "mensaje": "No existen registros de becas con contratos pendientes de generar por los estudiantes para generar el reporte"})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al generar la consulta."})

        elif action == 'verificar_becas_documentacion_rechazada':
            try:
                becas1 = BecaAsignacion.objects.filter(
                    Q(solicitud__inscripcion__persona__personadocumentopersonal__estadocedula=3) |
                    Q(solicitud__inscripcion__persona__personadocumentopersonal__estadopapeleta=3) |
                    Q(solicitud__inscripcion__persona__personadocumentopersonal__estadocedularepresentantesol=3) |
                    Q(solicitud__inscripcion__persona__personadocumentopersonal__estadopapeletarepresentantesol=3),
                    status=True, solicitud__periodo=periodo, tipo=1, cargadocumento=True).exclude(
                    solicitud__becatipo_id=16)

                becas2 = BecaAsignacion.objects.filter(
                    Q(solicitud__inscripcion__persona__personadocumentopersonal__estadocedula=3) |
                    Q(solicitud__inscripcion__persona__personadocumentopersonal__estadopapeleta=3) |
                    Q(solicitud__inscripcion__persona__personadocumentopersonal__estadocedularepresentantesol=3) |
                    Q(solicitud__inscripcion__persona__personadocumentopersonal__estadopapeletarepresentantesol=3) |
                    Q(solicitud__inscripcion__persona__personadocumentopersonal__estadoactagrado=3),
                    status=True, solicitud__periodo=periodo, tipo=1, solicitud__becatipo_id=16, cargadocumento=True)

                becas = becas1 | becas2

                if becas:
                    return JsonResponse({"result": "ok"})
                else:
                    return JsonResponse({"result": "bad",
                                         "mensaje": "No existen registros de becas con documentación rechazada para generar el reporte"})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al generar la consulta."})

        elif action == 'verificar_becas_solicitud_pago':
            try:
                if datetime.strptime(request.POST['desde'], '%d-%m-%Y') <= datetime.strptime(request.POST['hasta'],'%d-%m-%Y'):
                    from datetime import time
                    desde = datetime.combine(convertir_fecha(request.POST['desde']), time.min)
                    hasta = datetime.combine(convertir_fecha(request.POST['hasta']), time.max)
                    tipoasignacion = int(request.POST['tipoasignacion'])
                    tipobeca = int(request.POST['tipobeca'])

                    becas = BecaAsignacion.objects.filter(status=True, solicitud__periodo=periodo,
                                                          solicitud__becatipo_id=tipobeca,
                                                          solicitudpagobecadetalle__asignacion__isnull=True,
                                                          tipo=tipoasignacion, solicitud__status=True,
                                                          solicitud__becaaceptada=2)

                    # if tipoasignacion == 2:
                    #     becas = becas.filter(fecha_creacion__range=(desde, hasta))

                    if becas:
                        if tipoasignacion == 1:
                            becas = becas.filter(solicitud__inscripcion__persona__personadocumentopersonal__status=True,
                                                # fecha_modificacion__range=(desde, hasta),
                                                 solicitud__inscripcion__persona__personadocumentopersonal__estadocedula=2,
                                                 # solicitud__inscripcion__persona__personadocumentopersonal__estadopapeleta=2,
                                                 # solicitud__inscripcion__persona__personadocumentopersonal__estadocedularepresentantesol=2,
                                                 # solicitud__inscripcion__persona__personadocumentopersonal__estadopapeletarepresentantesol=2,
                                                 solicitud__inscripcion__persona__cuentabancariapersona__status=True,
                                                 solicitud__inscripcion__persona__cuentabancariapersona__estadorevision=2,
                                                 solicitud__inscripcion__persona__cuentabancariapersona__activapago=True,
                                                 solicitud__archivoactacompromiso__isnull=False
                                                 # estadorevisioncontrato=2
                                                 ).distinct()
                            idbs = [beca.pk for beca in becas if beca.solicitud.cumple_todos_documentos_requeridos_aprobados()]
                            becas = becas.filter(id__in=idbs)
                        becaperiodo = periodo.becaperiodo_set.filter(status=True).first()
                        if becaperiodo is not None:
                            limite = becaperiodo.limitebecados
                            numbecadospago = SolicitudPagoBeca.objects.filter(status=True, periodo=periodo).values_list('id', flat=True).count()
                            if numbecadospago > limite:
                                becas = becas[:limite]
                        if becas:
                            data['becas'] = becas.order_by('solicitud__inscripcion__persona__apellido1',
                                                           'solicitud__inscripcion__persona__apellido2',
                                                           'solicitud__inscripcion__persona__nombres')

                            data['periodoaplica'] = periodo.nombre
                            data['tipoasignacion'] = 'NUEVA' if tipoasignacion == 1 else 'RENOVACIÓN'
                            data['tipobeca'] = BecaTipo.objects.get(pk=tipobeca)
                            data['totalbeneficiarios'] = becas.count()
                            data['totalpagar'] = becas.aggregate(valor=Sum('montobeneficio'))['valor']
                            template = get_template("adm_becas/solicitudpago.html")
                            json_content = template.render(data, request=request)
                            return JsonResponse({"result": "ok", 'data': json_content})
                        else:
                            return JsonResponse({"result": "bad",
                                                 "mensaje": "No existen registros de becas con toda la documentación validada para generar el reporte"})
                    else:
                        return JsonResponse(
                            {"result": "bad", "mensaje": "No existen registros de becas para generar el reporte"})
                else:
                    return JsonResponse(
                        {"result": "bad", "mensaje": "La fecha desde debe ser menor o igual a la fecha hasta"})

            except Exception as ex:
                msg = 'Error on line {} {}'.format(sys.exc_info()[-1].tb_lineno, ex.__str__())
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al generar la consulta  %s."%(msg)})

        elif action == 'generar_solicitud_pago':
            try:
                from datetime import time
                desde = datetime.combine(convertir_fecha(request.POST['desde']), time.min)
                hasta = datetime.combine(convertir_fecha(request.POST['hasta']), time.max)
                tipoasignacion = int(request.POST['tipoasignacion'])
                tipobeca = int(request.POST['tipobeca'])

                becas = BecaAsignacion.objects.filter(status=True, solicitud__periodo=periodo,
                                                      solicitud__becatipo_id=tipobeca,
                                                      solicitudpagobecadetalle__asignacion__isnull=True,
                                                      tipo=tipoasignacion, solicitud__status=True,
                                                      solicitud__becaaceptada=2)

                # if tipoasignacion == 2:
                #     becas = becas.filter(fecha_creacion__range=(desde, hasta))

                if becas:
                    if tipoasignacion == 1:
                        becas = becas.filter(
                                                # fecha_modificacion__range=(desde, hasta),
                                                 solicitud__inscripcion__persona__personadocumentopersonal__status=True,
                                                 solicitud__inscripcion__persona__personadocumentopersonal__estadocedula=2,
                                                 # solicitud__inscripcion__persona__personadocumentopersonal__estadopapeleta=2,
                                                 # solicitud__inscripcion__persona__personadocumentopersonal__estadocedularepresentantesol=2,
                                                 # solicitud__inscripcion__persona__personadocumentopersonal__estadopapeletarepresentantesol=2,
                                                 solicitud__inscripcion__persona__cuentabancariapersona__status=True,
                                                 solicitud__inscripcion__persona__cuentabancariapersona__estadorevision=2,
                                                 solicitud__inscripcion__persona__cuentabancariapersona__activapago=True,
                                                 solicitud__archivoactacompromiso__isnull=False).distinct()
                        idbs = [beca.pk for beca in becas if beca.solicitud.cumple_todos_documentos_requeridos_aprobados()]
                        becas = becas.filter(id__in=idbs)
                    becaperiodo = periodo.becaperiodo_set.filter(status=True).first()
                    if becaperiodo is not None:
                        limite = becaperiodo.limitebecados
                        numbecadospago = SolicitudPagoBeca.objects.filter(status=True, periodo=periodo).values_list('id', flat=True).count()
                        if numbecadospago > limite:
                            becas = becas[:limite]
                    if becas:
                        # becas = becas.filter(solicitud__inscripcion__persona__cuentabancariapersona__estadorevision=2)
                        # if becas:
                        secuencia = secuencia_solicitud_pago_beca(periodo)
                        cantidadbeneficiarios = becas.count()
                        montopago = becas.aggregate(valor=Sum('montobeneficio'))['valor']
                        solicitudpago = SolicitudPagoBeca(fecha=datetime.now().date(),
                                                          cantidadbenef=cantidadbeneficiarios,
                                                          montopago=montopago,
                                                          pagado=False,
                                                          becatipo_id=tipobeca,
                                                          tipoasignacion=tipoasignacion,
                                                          periodo=periodo,
                                                          numerosolicitud=secuencia)
                        solicitudpago.save(request)
                        for beca in becas:
                            if not SolicitudPagoBecaDetalle.objects.filter(asignacion=beca,status=True).exists():
                                detallepago = SolicitudPagoBecaDetalle(solicitudpago=solicitudpago,
                                                                       asignacion=beca,
                                                                       monto=beca.montobeneficio,
                                                                       pagado=False,
                                                                       acreditado=False)
                                detallepago.save(request)

                                recorrido = BecaSolicitudRecorrido(solicitud=beca.solicitud,
                                                                   observacion='SOLICITUD DE PAGO REALIZADA',
                                                                   estado=28,
                                                                   fecha=datetime.now().date()
                                                                   )
                                recorrido.save()

                        totalsidetalle=SolicitudPagoBecaDetalle.objects.filter(solicitudpago=solicitudpago,status=True)
                        solicitudpago.cantidadbenef=len(totalsidetalle)
                        solicitudpago.montopago=totalsidetalle.aggregate(valor=Sum('monto'))['valor']
                        solicitudpago.save(request)

                        log(u'Agregó solicitud de pago de beca: %s' % solicitudpago, request, "add")
                        return JsonResponse({"result": "ok", "ids": solicitudpago.id})
                        # else:
                        #     return JsonResponse({"result": "bad", "mensaje": "Los registros de becas disponibles para la solicitud de pago no tienen las cuentas bancarias validadas por Tesorería"})
                    else:
                        return JsonResponse({"result": "bad",
                                             "mensaje": "No existen registros de becas con toda la documentación validada para generar el reporte"})
                else:
                    return JsonResponse(
                        {"result": "bad", "mensaje": "No existen registros de becas para generar el reporte"})

            except Exception as ex:
                msg = 'Error on line {} {}'.format(sys.exc_info()[-1].tb_lineno, ex.__str__())
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al generar la consulta  %s."%(msg)})

        elif action == 'aprobar_rechazar_solicitud':
            try:
                LIMITE_BECAS_PRESENCIAL_SEMIPRESENCIAL = 1150  # 1150
                LIMITE_BECAS_ENLINEA = 450  # 450

                idsolicitud = int(encrypt(request.POST['id']))
                tipoaccion = request.POST['tipoaccion']
                estado = 2 if tipoaccion == 'A' else 3
                accionlog = 'Aprobó' if tipoaccion == 'A' else 'Rechazó'

                solicitud = BecaSolicitud.objects.get(pk=idsolicitud)

                if tipoaccion == 'A':
                    solicitudes = BecaSolicitud.objects.filter(periodo=periodo, status=True, estado=2).exclude(
                        becaaceptada=3)
                    totalpresemi = solicitudes.filter(inscripcion__modalidad_id__in=[1, 2]).count()
                    totalenlinea = solicitudes.filter(inscripcion__modalidad_id=3).count()

                    if solicitud.inscripcion.modalidad.id in [1, 2]:
                        if totalpresemi == LIMITE_BECAS_PRESENCIAL_SEMIPRESENCIAL:
                            return JsonResponse({"result": "bad",
                                                 "mensaje": u"El cupo para asignación de becas para TABLET + PLAN DE DATOS se agotó."})
                    else:
                        if totalenlinea == LIMITE_BECAS_ENLINEA:
                            return JsonResponse({"result": "bad",
                                                 "mensaje": u"El cupo para asignación de becas para PLAN DE DATOS se agotó."})

                # solicitud.becaasignada = 1 if estado == 2 else 2
                solicitud.estado = estado
                solicitud.becaaceptada = 1
                solicitud.save(request)

                recorrido = BecaSolicitudRecorrido(solicitud=solicitud,
                                                   estado=estado,
                                                   fecha=datetime.now().date())
                recorrido.save(request)

                log(u'%s una solicitud de beca: %s - %s' % (accionlog, solicitud, persona), request, "edit")

                # Envio de e-mail de notificación al solicitante
                tituloemail = "Solicitud de Beca " + "Aprobada" if estado == 2 else "Rechazada"

                archivos = []
                observaciones = []

                observaciondetalle = ""
                if estado == 3:
                    rechazados = solicitud.becadetallesolicitud_set.values('requisito__nombre', 'archivo',
                                                                           'observacion', 'estado').filter(estado=3,
                                                                                                           requisito_id__in=[
                                                                                                               9, 10,
                                                                                                               15, 16,
                                                                                                               30,
                                                                                                               33]).order_by(
                        'id')
                    for req in rechazados:
                        archivos.append(req['archivo'])
                        observaciones.append(req['observacion'])

                        if observaciondetalle == '':
                            observaciondetalle = "<strong>- " + req['requisito__nombre'] + "</strong>:<br>" + req[
                                'observacion']
                        else:
                            observaciondetalle = observaciondetalle + "<br><strong>- " + req[
                                'requisito__nombre'] + "</strong>:<br>" + req['observacion']

                recorrido.observacion = observaciondetalle if estado == 3 else 'APROBADO'
                recorrido.save(request)

                tipobeca = solicitud.becatipo.nombre.upper()
                send_html_mail(tituloemail,
                               "emails/notificarestadosolicitudbeca.html",
                               {'sistema': u'SGA - UNEMI',
                                'fase': 'AR',
                                'tipobeca': tipobeca,
                                'fecha': datetime.now().date(),
                                'hora': datetime.now().time(),
                                'saludo': 'Estimada' if solicitud.inscripcion.persona.sexo_id == 1 else 'Estimado',
                                'estado': 'APROBADA' if estado == 2 else "RECHAZADA",
                                'estudiante': solicitud.inscripcion.persona.nombre_completo_inverso(),
                                'autoridad2': '',
                                'observaciones': observaciones,
                                't': miinstitucion()
                                },
                               solicitud.inscripcion.persona.lista_emails_envio(),
                               [],
                               archivos,
                               cuenta=variable_valor('CUENTAS_CORREOS')[0]
                               )

                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al procesar el Registro."})

        elif action == 'asignarsolicitud':
            try:
                inscripcion = Inscripcion.objects.get(id=int(request.POST['id']))
                becatipo = BecaTipo.objects.get(pk=int(request.POST['tipo']))
                if not BecaSolicitud.objects.values('id').filter(inscripcion=inscripcion, periodo=periodo).exists():
                    observacion = 'solicitado por: ' + persona.nombre_completo_inverso()
                    becasolicitud = BecaSolicitud(inscripcion=inscripcion,
                                                  becatipo=becatipo,
                                                  periodo=periodo,
                                                  estado=2, observacion=observacion)
                    becasolicitud.save(request)
                    if not inscripcion.matricula().nivelmalla.id == 1:
                        for x in BecaRequisitos.objects.filter(status=True):
                            # if x.id== 1 or x.id== 3 or x.id== 5:
                            #     cumple=True
                            # else:
                            #     cumple=False
                            becasolicituddetalle = BecaDetalleSolicitud(solicitud=becasolicitud,
                                                                        requisito=x,
                                                                        cumple=True,
                                                                        estado=2)
                            becasolicituddetalle.save(request)
                    else:
                        becasolicituddetalle = BecaDetalleSolicitud(solicitud=becasolicitud,
                                                                    requisito_id=1,
                                                                    cumple=True,
                                                                    estado=4)
                        becasolicituddetalle.save(request)
                        # becasolicituddetalle = BecaDetalleSolicitud(solicitud=becasolicitud,
                        #                                             requisito_id=7,
                        #                                             cumple=True,
                        #                                             estado=4)
                        # becasolicituddetalle.save(request)
                    log(u'Asigno solicitud de un estudiante a beca: %s' % becasolicitud.inscripcion, request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                    return JsonResponse({"result": "bad", "mensaje": u"Inscripcion ya asignada."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al Asignar, Intentelo mas tarde."})

        elif action == 'asignarbeca':
            try:
                idsolicitud = int(request.POST['idsolicitud'])
                f = BecaAsignacionForm(request.POST)
                if f.is_valid():
                    montomensual = null_to_decimal(f.cleaned_data['montomensual'], 2)
                    montobeneficio = null_to_decimal(f.cleaned_data['montobeneficio'], 2)
                    if montomensual <= 0 or montobeneficio <= 0:
                        return JsonResponse({"result": "bad", "mensaje": u"Ingrese montos mayores a cero."})
                    cantidadmeses = int(f.cleaned_data['cantidadmeses'])
                    total = null_to_decimal(montomensual * cantidadmeses, 2)
                    if montobeneficio != total:
                        return JsonResponse({"result": "bad", "mensaje": u"montos no cuadran."})
                    if not BecaAsignacion.objects.values('id').filter(solicitud__id=idsolicitud,
                                                                      solicitud__periodo=periodo).exists():
                        if BecaSolicitud.objects.values('id').filter(id=idsolicitud, periodo=periodo).exists():
                            sol = BecaSolicitud.objects.get(id=idsolicitud)
                            becaasignacion = BecaAsignacion(solicitud=sol,
                                                            montomensual=f.cleaned_data['montomensual'],
                                                            cantidadmeses=f.cleaned_data['cantidadmeses'],
                                                            montobeneficio=f.cleaned_data['montobeneficio'],
                                                            fecha=datetime.now().date(),
                                                            tipo=int(f.cleaned_data['tipo']),
                                                            grupopago=int(f.cleaned_data['grupopago']) if
                                                            f.cleaned_data['grupopago'] else None,
                                                            activo=True)
                            becaasignacion.save(request)
                            log(u'Asigno estudiante a beca: %s' % becaasignacion.solicitud, request, "add")
                        return JsonResponse({"result": "ok"})
                    else:
                        return JsonResponse({"result": "bad", "mensaje": u"Inscripcion ya asignada."})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'aprobarsolicitud':
            try:
                id = int(request.POST['id'])
                observacion = request.POST['observacion']
                estado = int(request.POST['estado'])
                if BecaSolicitud.objects.values('id').filter(id=id, periodo=periodo).exists():
                    becasolicitud = BecaSolicitud.objects.get(pk=id)
                    becasolicitud.estado = estado
                    becasolicitud.observacion = observacion
                    becasolicitud.save(request)
                    # for x in becasolicitud.becadetallesolicitud_set.all():
                    #     x.estado=estado
                    #     x.save(request)
                    log(u'Asigno solicitud de un estudiante a beca: %s' % becasolicitud.inscripcion, request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                    return JsonResponse({"result": "bad", "mensaje": u"No existe."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al Asignar, Intentelo mas tarde."})

        elif action == 'aprobarrechazararchivo':
            try:
                detalle = BecaDetalleSolicitud.objects.get(id=int(request.POST['idevidencia']),
                                                           solicitud_id=int(request.POST['idc']))

                f = BecaAprobarArchivoForm(request.POST, request.FILES)

                if 'archivo' in request.FILES:
                    arch = request.FILES['archivo']
                    extension = arch._name.split('.')
                    tam = len(extension)
                    exte = extension[tam - 1]
                    if arch.size > 6291456:
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"Error, el tamaño del archivo es mayor a 6 Mb."})
                    if not exte.lower() == 'pdf':
                        return JsonResponse({"result": "bad", "mensaje": u"Solo se permiten archivos .pdf"})

                if f.is_valid():
                    estado = int(f.cleaned_data['estado'])
                    detalle.cumple = True if estado == 2 else False
                    detalle.estado = f.cleaned_data['estado']
                    detalle.observacion = f.cleaned_data['observacion'].upper()
                    detalle.fechaaprueba = datetime.now()
                    detalle.personaaprueba = persona
                    if 'archivo' in request.FILES:
                        newfile = request.FILES['archivo']
                        newfile._name = generar_nombre("requisito", newfile._name)
                        detalle.archivo = newfile

                    detalle.save(request)
                    log(u'Adiciono pruebas de evidencia en becas: %s [%s]' % (
                        detalle.id, detalle.solicitud.inscripcion), request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'aprobarrechazarutilizacion':
            try:
                detalle = BecaDetalleUtilizacion.objects.get(id=int(request.POST['idevidencia']))
                f = BecaAprobarArchivoUtilizacionForm(request.POST)
                if f.is_valid():
                    detalle.estado = f.cleaned_data['estado']
                    detalle.observacion = f.cleaned_data['observacion']
                    detalle.fechaaprueba = datetime.now()
                    detalle.personaaprueba = persona
                    detalle.save(request)
                    log(u'Aprobo o Rechazo evidencia de utilizacion de beca: %s - %s - %s -%s' % (
                        detalle.estado, detalle.personaaprueba, detalle.fechaaprueba, detalle.archivo.url), request,
                        "edit")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'anularbeca':
            try:
                beca = BecaAsignacion.objects.get(pk=int(request.POST['id']))
                solicitud = BecaSolicitud.objects.get(pk=beca.solicitud.id)

                if beca.estadobeca == 3:
                    return JsonResponse({"result": "bad", "mensaje": u"La beca ya ha sido anulada por otro usuario."})

                f = BecaSolicitudAnulacionForm(request.POST, request.FILES)

                # if not 'archivo' in request.FILES:
                #     return JsonResponse(
                #         {"result": "bad", "mensaje": u"Atención, debe seleccionar un archivo en formato PDF."})



                if not f.is_valid():
                    for k, v in f.errors.items():
                        raise NameError(v[0])

                if 'archivo' in request.FILES:
                    arch = request.FILES['archivo']
                    extension = arch._name.split('.')
                    tam = len(extension)
                    exte = extension[tam - 1]
                    if arch.size > 4194304:
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"Error, el tamaño del archivo es mayor a 4 Mb."})
                    if not exte.lower() == 'pdf':
                        return JsonResponse({"result": "bad", "mensaje": u"Solo se permiten archivos .pdf"})
                    newfile = request.FILES['archivo']
                    newfile._name = generar_nombre("becaanul", newfile._name)
                    beca.archivoanulacion = newfile
                beca.estadobeca = 3
                beca.status = False
                beca.save(request)

                solicitud.becaaceptada = 3
                solicitud.save(request)

                detalle = BecaSolicitudRecorrido(solicitud=solicitud,
                                                 estado=7,
                                                 observacion=f.cleaned_data['observacion'].strip().upper(),
                                                 fecha=datetime.now().date()
                                                 )
                detalle.save(request)
                if (eBecaPeriodo := solicitud.periodo.becaperiodo_set.filter(status=True).first()) is not None:
                    if eBecaPeriodo.fechainiciosolicitud >= datetime.now() and eBecaPeriodo.fechafinsolicitud <= datetime.now():
                        # Envio de e-mail de notificación al solicitante
                        tituloemail = "Beca Anulada"

                        send_html_mail(tituloemail,
                                       "emails/notificaranulacionbeca.html",
                                       {'sistema': u'SGA - UNEMI',
                                        'fecha': datetime.now().date(),
                                        'hora': datetime.now().time(),
                                        'saludo': 'Estimada' if solicitud.inscripcion.persona.sexo_id == 1 else 'Estimado',
                                        'estudiante': solicitud.inscripcion.persona.nombre_completo_inverso(),
                                        'periodo': solicitud.periodo.__str__(),
                                        'autoridad2': '',
                                        'observaciones': f.cleaned_data['observacion'].strip().upper(),
                                        't': miinstitucion()
                                        },
                                       solicitud.inscripcion.persona.lista_emails_envio(),
                                       [],
                                       cuenta=variable_valor('CUENTAS_CORREOS')[0]
                                       )

                log(u'Anuló beca: %s  - %s' % (persona, solicitud), request, "edit")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'actualizarpagado':
            try:
                f = ImportarArchivoPagoBecaXLSForm(request.POST, request.FILES)

                if not 'archivo' in request.FILES:
                    return JsonResponse(
                        {"result": "bad", "mensaje": u"Atención, debe seleccionar un archivo en formato XLSX."})

                if 'archivo' in request.FILES:
                    arch = request.FILES['archivo']
                    extension = arch._name.split('.')
                    tam = len(extension)
                    exte = extension[tam - 1]
                    if arch.size > 12582912:
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"Error, el tamaño del archivo es mayor a 12 Mb."})
                    if not exte.lower() == 'xlsx':
                        return JsonResponse({"result": "bad", "mensaje": u"Solo se permiten archivos .xls"})

                if f.is_valid():
                    liborigen = load_workbook(request.FILES['archivo'])
                    sheet = liborigen._sheets[0]
                    fila = 1
                    # VERIFICAR POSIBLES NOVEDADES
                    noexiste = pagado = False

                    for rowx in sheet.rows:
                        if fila >= 2:
                            print(fila - 1)
                            cedula = str(sheet.cell(row=fila, column=1).value).strip()
                            if len(cedula) == 9:
                                cedula = '0' + cedula

                            nombres = str(sheet.cell(row=fila, column=2).value)
                            fechapago = str(sheet.cell(row=fila, column=4).value)
                            monto = str(sheet.cell(row=fila, column=3).value)
                            if not SolicitudPagoBecaDetalle.objects.values("id").filter(
                                    asignacion__solicitud__periodo=periodo,
                                    asignacion__solicitud__inscripcion__persona__cedula=cedula).exists():
                                noexiste = True
                                break
                            elif SolicitudPagoBecaDetalle.objects.values("id").filter(
                                    asignacion__solicitud__periodo=periodo,
                                    asignacion__solicitud__inscripcion__persona__cedula=cedula, pagado=True).exists():
                                pagado = True
                                break

                        fila += 1

                    if pagado is False and noexiste is False:
                        newfile = request.FILES['archivo']
                        newfile._name = generar_nombre("pagobeca_", newfile._name)

                        archivopago = ArchivoPagoBeca(nombre='PAGO DE BECA',
                                                      fecha=datetime.now().date(),
                                                      archivo=newfile,
                                                      tipo=1)
                        archivopago.save(request)

                        fila = 1
                        for rowx in sheet.rows:
                            if fila >= 2:
                                cedula = str(sheet.cell(row=fila, column=1).value).strip()
                                if len(cedula) == 9:
                                    cedula = '0' + cedula

                                nombres = str(sheet.cell(row=fila, column=2).value)
                                if not isinstance(sheet.cell(row=fila, column=4).value, datetime):
                                    raise NameError('Formato de documento cargado incorrecto: Columna 4 o D, tine que contener valores de tipo fecha.')
                                fechapago = str(sheet.cell(row=fila, column=4).value)
                                monto = str(sheet.cell(row=fila, column=3).value)

                                beca = BecaAsignacion.objects.get(solicitud__periodo=periodo,
                                                                  solicitud__inscripcion__persona__cedula=cedula,status=True)
                                solicitudpago = beca.solicitudpagobecadetalle_set.filter(status=True)[0]

                                if solicitudpago.pagado is False:
                                    solicitudpago.pagado = True
                                    fpago = datetime.strptime(fechapago[:10], '%Y-%m-%d').date()
                                    solicitudpago.fechapago = fpago
                                    solicitudpago.personapago = persona
                                    solicitudpago.save(request)

                                    observacion = 'PAGADO POR UNEMI. EL PAGO SE ENCUENTRA PENDIENTE DE ACREDITAR POR PARTE DEL MINISTERIO DE ECONOMÍA Y FINANZAS'

                                    recorrido = BecaSolicitudRecorrido(solicitud=beca.solicitud,
                                                                       observacion=observacion,
                                                                       estado=29,
                                                                       fecha=fpago
                                                                       )
                                    recorrido.save(request)

                                    log(u'Actualizó beca a estado PAGADO: %s  - %s' % (persona, beca.solicitud),
                                        request, "edit")

                            fila += 1

                        return JsonResponse({"result": "ok"})
                    else:
                        if noexiste:
                            return JsonResponse({"result": "bad",
                                                 "mensaje": u"El registro con cédula # %s no consta en la solicitud de pago." % cedula})
                        else:
                            return JsonResponse({"result": "bad",
                                                 "mensaje": u"El registro con cédula # %s ya tiene estado PAGADO." % cedula})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": f"Error:{ex}"})

        elif action == 'actualizaracreditado':
            try:
                f = ImportarArchivoPagoBecaXLSForm(request.POST, request.FILES)

                if not 'archivo' in request.FILES:
                    return JsonResponse(
                        {"result": "bad", "mensaje": u"Atención, debe seleccionar un archivo en formato XLSX."})

                if 'archivo' in request.FILES:
                    arch = request.FILES['archivo']
                    extension = arch._name.split('.')
                    tam = len(extension)
                    exte = extension[tam - 1]
                    if arch.size > 12582912:
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"Error, el tamaño del archivo es mayor a 12 Mb."})
                    if not exte.lower() == 'xlsx':
                        return JsonResponse({"result": "bad", "mensaje": u"Solo se permiten archivos .xls"})

                if f.is_valid():
                    liborigen = load_workbook(request.FILES['archivo'])
                    sheet = liborigen._sheets[0]
                    fila = 1
                    # VERIFICAR POSIBLES NOVEDADES
                    noexiste = acreditado = False

                    for rowx in sheet.rows:
                        if fila >= 2:
                            print(fila - 1)
                            cedula = str(sheet.cell(row=fila, column=1).value).strip()
                            print(cedula)
                            if len(cedula) == 9:
                                cedula = '0' + cedula

                            nombres = str(sheet.cell(row=fila, column=2).value)
                            fechapago = str(sheet.cell(row=fila, column=4).value)
                            monto = str(sheet.cell(row=fila, column=3).value)
                            if not SolicitudPagoBecaDetalle.objects.values("id").filter(
                                    asignacion__solicitud__periodo=periodo,
                                    asignacion__solicitud__inscripcion__persona__cedula=cedula,
                                    pagado=True).exists():
                                noexiste = True
                                break
                            elif SolicitudPagoBecaDetalle.objects.values("id").filter(
                                    asignacion__solicitud__periodo=periodo,
                                    asignacion__solicitud__inscripcion__persona__cedula=cedula,
                                    pagado=True, acreditado=True).exists():
                                acreditado = True
                                break

                        fila += 1

                        if fila >= 28:
                            print("Hy")

                    if acreditado is False and noexiste is False:
                        newfile = request.FILES['archivo']
                        newfile._name = generar_nombre("acreditabeca_", newfile._name)

                        archivopago = ArchivoPagoBeca(nombre='ACREDITACIÓN DE BECA',
                                                      fecha=datetime.now().date(),
                                                      archivo=newfile,
                                                      tipo=2)
                        archivopago.save(request)

                        fila = 1
                        for rowx in sheet.rows:
                            if fila >= 2:
                                cedula = str(sheet.cell(row=fila, column=1).value).strip()
                                if len(cedula) == 9:
                                    cedula = '0' + cedula

                                nombres = str(sheet.cell(row=fila, column=2).value)
                                fechaacredita = str(sheet.cell(row=fila, column=4).value)
                                monto = str(sheet.cell(row=fila, column=3).value)

                                beca = BecaAsignacion.objects.get(solicitud__periodo=periodo,
                                                                  solicitud__inscripcion__persona__cedula=cedula,status=True)
                                solicitudpago = beca.solicitudpagobecadetalle_set.filter(status=True)[0]

                                if solicitudpago.acreditado is False:
                                    solicitudpago.acreditado = True
                                    facredita = datetime.strptime(fechaacredita[:10], '%Y-%m-%d').date()
                                    solicitudpago.fechaacredita = facredita
                                    solicitudpago.personaacredita = persona
                                    solicitudpago.save(request)

                                    observacion = 'ACREDITADO POR MINISTERIO DE ECONOMÍA Y FINANZAS A TRAVÉS DE TRANSFERENCIA BANCARIA'

                                    recorrido = BecaSolicitudRecorrido(solicitud=beca.solicitud,
                                                                       observacion=observacion,
                                                                       estado=30,
                                                                       fecha=facredita
                                                                       )
                                    recorrido.save(request)

                                    log(u'Actualizó beca a estado ACREDITADO: %s  - %s' % (persona, beca.solicitud),
                                        request, "edit")

                            fila += 1

                        return JsonResponse({"result": "ok"})
                    else:
                        if noexiste:
                            return JsonResponse({"result": "bad",
                                                 "mensaje": u"El registro con cédula # %s no consta en la solicitud de pago." % (
                                                             cedula + " - " + str(fila))})
                        else:
                            return JsonResponse({"result": "bad",
                                                 "mensaje": u"El registro con cédula # %s ya tiene estado ACREDITADO." % cedula})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'rechazarsolicitudonline':
            try:
                solicitud = BecaSolicitud.objects.get(pk=int(request.POST['id']))

                if solicitud.estado == 3:
                    return JsonResponse(
                        {"result": "bad", "mensaje": u"La solicitud ya ha sido rechazada por otro usuario."})

                f = BecaSolicitudRechazaOnLineForm(request.POST, request.FILES)

                # if not 'archivo' in request.FILES:
                #     return JsonResponse({"result": "bad", "mensaje": u"Atención, debe seleccionar un archivo en formato PDF."})

                # if 'archivo' in request.FILES:
                #     arch = request.FILES['archivo']
                #     extension = arch._name.split('.')
                #     tam = len(extension)
                #     exte = extension[tam - 1]
                #     if arch.size > 4194304:
                #         return JsonResponse({"result": "bad", "mensaje": u"Error, el tamaño del archivo es mayor a 4 Mb."})
                #     if not exte.lower() == 'pdf':
                #         return JsonResponse({"result": "bad", "mensaje": u"Solo se permiten archivos .pdf"})

                if f.is_valid():
                    solicitud.estado = 3
                    solicitud.observacion = f.cleaned_data['observacion'].strip().upper()

                    # newfile = request.FILES['archivo']
                    # newfile._name = generar_nombre("solbecaanul", newfile._name)
                    # solicitud.archivoanulacion = newfile

                    solicitud.save(request)

                    recorrido = BecaSolicitudRecorrido(
                        solicitud=solicitud,
                        observacion=f.cleaned_data['observacion'].strip().upper(),
                        estado=3,
                        fecha=datetime.now().date()
                    )
                    recorrido.save(request)

                    tituloemail = "Solicitud de Beca Rechazada"

                    observaciones = [f.cleaned_data['observacion'].strip().upper()]
                    send_html_mail(tituloemail,
                                   "emails/notificarestadosolicitudbeca.html",
                                   {'sistema': u'SGA - UNEMI',
                                    'fase': 'AR',
                                    'fecha': datetime.now().date(),
                                    'hora': datetime.now().time(),
                                    'observaciones': observaciones,
                                    'saludo': 'Estimada' if solicitud.inscripcion.persona.sexo_id == 1 else 'Estimado',
                                    'estado': 'RECHAZADA',
                                    'estudiante': solicitud.inscripcion.persona.nombre_completo_inverso(),
                                    'autoridad2': '',
                                    't': miinstitucion()
                                    },
                                   solicitud.inscripcion.persona.lista_emails_envio(),
                                   [],
                                   cuenta=variable_valor('CUENTAS_CORREOS')[0]
                                   )

                    log(u'Rechazó la solicitud de beca: %s  - %s' % (persona, solicitud), request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'validarcedula':
            try:
                beca = BecaAsignacion.objects.get(pk=int(request.POST['id']))
                estudiante = beca.solicitud.inscripcion.persona
                representante = estudiante.personaextension_set.all()[0]
                if not BecaPeriodo.objects.filter(periodo=beca.solicitud.periodo, vigente=True):
                    return JsonResponse(
                        {"result": "bad", "mensaje": u"Error, no se encontro el periodo de beca configurado"})
                becaperiodo = BecaPeriodo.objects.filter(periodo=beca.solicitud.periodo, vigente=True).first()
                # if not becaperiodo.puede_revisar_validar_documentos():
                #     return JsonResponse({"result": "bad", "mensaje":"Se completo el límite de becados del período %s"%(becaperiodo)})
                fechainicioimprimircontrato = becaperiodo.fechainicioimprimircontrato
                # resp = validarcedula(request.POST['numerocedularepresentante'].strip())
                # if resp != 'Ok':
                #     return JsonResponse({"result": "bad", "mensaje": resp})
                lista_correos = beca.solicitud.inscripcion.persona.lista_emails_envio(),
                if DEBUG:
                    lista_correos = ['atorrese@unemi.edu.ec',]
                # if request.POST['numerocedularepresentante'].strip() == estudiante.cedula:
                #     return JsonResponse({"result": "bad", "mensaje": "El número de cédula del representante solidario debe ser distinto al del estudiante"})

                # representante.cedularepsolidario = request.POST['numerocedularepresentante'].strip()
                # representante.nombresrepsolidario = request.POST['nombrerepresentante'].strip().upper()
                # representante.apellido1repsolidario = request.POST['apellido1representante'].strip().upper()
                # representante.apellido2repsolidario = request.POST['apellido2representante'].strip().upper()
                # representante.save(request)

                documentos = beca.solicitud.inscripcion.persona.documentos_personales()

                documentos.estadocedula = int(request.POST['estadocedula1'])
                #documentos.estadopapeleta = int(request.POST['estadocert1'])
                #documentos.estadocedularepresentantesol = int(request.POST['estadocedula2'])
                #documentos.estadopapeletarepresentantesol = int(request.POST['estadocert2'])
                if beca.solicitud.becatipo.id == 22 and beca.solicitud.inscripcion.persona.ecuatoriano_vive_exterior():
                    documentos.estadoserviciosbasico = int(request.POST['estadoserviciobasico'])
                documentos.observacion = request.POST['observacion'].strip().upper() if 'observacion' in request.POST else ''
                documentos.save(request)

                estadoarchivoraza = 0
                if beca.solicitud.becatipo.id == 21:
                    if beca.solicitud.inscripcion.persona.perfilinscripcion_set.filter(status=True).exists():
                        perilinscripcion = beca.solicitud.inscripcion.persona.perfilinscripcion_set.filter(
                            status=True).first()
                        estadoarchivoraza = int(request.POST['estadoarchivoraza'])
                        perilinscripcion.estadoarchivoraza = estadoarchivoraza
                        perilinscripcion.save(request)

                if documentos.estadocedula == 3 or documentos.estadocedularepresentantesol == 3 or documentos.estadopapeleta == 3 or documentos.estadopapeletarepresentantesol == 3 or documentos.estadoserviciosbasico == 3 or estadoarchivoraza == 3:
                    tituloemail = "Novedades con Documentación"
                    mensaje = ""
                    if documentos.estadocedula == 3:
                        mensaje = " se presentaron novedades con la revisión del archivo de su cédula de identidad"

                    # if documentos.estadopapeleta == 3:
                    #     if mensaje == "":
                    #         mensaje = " se presentaron novedades con la revisión del archivo de su certificado de votación"
                    #     else:
                    #         mensaje = mensaje + ", con el archivo de su certificado de votación"

                    if documentos.estadoserviciosbasico == 3:
                        if mensaje == "":
                            mensaje = " se presentaron novedades con la revisión del archivo de servicios básicos"
                        else:
                            mensaje = mensaje + ", con el archivo de servicios básicos"

                    if estadoarchivoraza == 3:
                        if mensaje == "":
                            mensaje = " se presentaron novedades con la revisión del archivo de la declaración juramentada"
                        else:
                            mensaje = mensaje + ", con el archivo de declaración juramentada"

                    # if documentos.estadocedularepresentantesol == 3:
                    #     if mensaje == "":
                    #         mensaje = " se presentaron novedades con la revisión del archivo de la cédula del representante solidario"
                    #     else:
                    #         mensaje = mensaje + ", el archivo de la cédula del representante solidario"
                    #
                    # if documentos.estadopapeletarepresentantesol == 3:
                    #     if mensaje == "":
                    #         mensaje = " se presentaron novedades con la revisión del archivo del certificado de votación del representante solidario"
                    #     else:
                    #         mensaje = mensaje + ", con el archivo del certificado de votación del representante solidario"

                    observaciones = documentos.observacion

                    recorrido = BecaSolicitudRecorrido(solicitud=beca.solicitud,
                                                       observacion=observaciones,
                                                       estado=12,
                                                       fecha=datetime.now().date())
                    recorrido.save(request)

                    send_html_mail(tituloemail,
                                   "emails/notificarrevisiondocumento.html",
                                   {'sistema': u'SGA - UNEMI',
                                    'mensaje': mensaje,
                                    'fecha': datetime.now().date(),
                                    'hora': datetime.now().time(),
                                    'observaciones': observaciones,
                                    'saludo': 'Estimada' if beca.solicitud.inscripcion.persona.sexo_id == 1 else 'Estimado',
                                    'estudiante': beca.solicitud.inscripcion.persona.nombre_completo_inverso(),
                                    'autoridad2': '',
                                    't': miinstitucion()
                                    },
                                   lista_correos,
                                   [],
                                   cuenta=variable_valor('CUENTAS_CORREOS')[0]
                                   )
                else:
                    recorrido = BecaSolicitudRecorrido(solicitud=beca.solicitud,
                                                       observacion="DOCUMENTACIÓN VALIDADA",
                                                       estado=11,
                                                       fecha=datetime.now().date())
                    recorrido.save(request)

                    tituloemail = "Beca - Documentación Validada"

                    if beca.solicitud.becatipo.id == 23:
                        mensaje = "los documentos correspondientes a la cédula y certificado de votación del estudiante y del representante solidario han sido Validados"
                        observaciones = "Usted deberá acceder al Módulo de Beca Estudiantil, descargar el contrato de ayuda económica, imprimirlo, recoger las firmas correspondientes y subirlo al sistema en formato PDF."
                    else:
                        if beca.solicitud.becatipo.id == 16:
                            mensaje = "los documentos correspondientes al estudiante y representante solidario han sido Validados"
                        else:
                            mensaje = "los documentos correspondientes a la cédula y certificado de votación del estudiante y del representante solidario han sido Validados"

                        observaciones = f"Usted deberá acceder al Módulo de Beca Estudiantil, descargar el contrato de beca (disponible el {leer_fecha_hora_letra_formato_fecha_hora(str(fechainicioimprimircontrato.date()), str(fechainicioimprimircontrato.time()), True)}), imprimirlo, recoger las firmas correspondientes y subirlo al sistema en formato PDF."

                    send_html_mail(tituloemail,
                                   "emails/notificarrevisiondocumento.html",
                                   {'sistema': u'SGA - UNEMI',
                                    'mensaje': mensaje,
                                    'fecha': datetime.now().date(),
                                    'hora': datetime.now().time(),
                                    'observaciones': observaciones,
                                    'saludo': 'Estimada' if beca.solicitud.inscripcion.persona.sexo_id == 1 else 'Estimado',
                                    'estudiante': beca.solicitud.inscripcion.persona.nombre_completo_inverso(),
                                    'autoridad2': '',
                                    't': miinstitucion()
                                    },
                                   lista_correos,
                                   [],
                                   cuenta=variable_valor('CUENTAS_CORREOS')[0]
                                   )

                log(u'Revisó documentación cédula de la beca: %s  - %s' % (persona, beca), request, "edit")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'validarcedulabpn':
            try:
                beca = BecaAsignacion.objects.get(pk=int(request.POST['id']))
                estudiante = beca.solicitud.inscripcion.persona
                representante = estudiante.personaextension_set.all()[0]
                resp = validarcedula(request.POST['numerocedularepresentante'].strip())
                if resp != 'Ok':
                    return JsonResponse({"result": "bad", "mensaje": resp})

                # if request.POST['numerocedularepresentante'].strip() == estudiante.cedula:
                #     return JsonResponse({"result": "bad", "mensaje": "El número de cédula del representante solidario debe ser distinto al del estudiante"})

                representante.cedularepsolidario = request.POST['numerocedularepresentante'].strip()
                representante.nombresrepsolidario = request.POST['nombrerepresentante'].strip().upper()
                representante.apellido1repsolidario = request.POST['apellido1representante'].strip().upper()
                representante.apellido2repsolidario = request.POST['apellido2representante'].strip().upper()
                representante.save(request)

                documentos = beca.solicitud.inscripcion.persona.documentos_personales()

                documentos.estadocedula = int(request.POST['estadocedula1'])
                documentos.estadopapeleta = int(request.POST['estadocert1'])
                documentos.estadocedularepresentantesol = int(request.POST['estadocedula2'])
                documentos.estadopapeletarepresentantesol = int(request.POST['estadocert2'])
                documentos.estadoactagrado = int(request.POST['estadoactagrado'])
                documentos.observacion = request.POST[
                    'observacion'].strip().upper() if 'observacion' in request.POST else ''
                documentos.save(request)

                if documentos.estadocedula == 3 or documentos.estadocedularepresentantesol == 3 or documentos.estadopapeleta == 3 or documentos.estadopapeletarepresentantesol == 3 or documentos.estadoactagrado == 3:
                    tituloemail = "Novedades con Documentación"
                    mensaje = ""
                    if documentos.estadocedula == 3:
                        mensaje = " se presentaron novedades con la revisión del archivo de su cédula de identidad"

                    if documentos.estadopapeleta == 3:
                        if mensaje == "":
                            mensaje = " se presentaron novedades con la revisión del archivo de su certificado de votación"
                        else:
                            mensaje = mensaje + ", con el archivo de su certificado de votación"

                    if documentos.estadoactagrado == 3:
                        if mensaje == "":
                            mensaje = " se presentaron novedades con la revisión del archivo de su acta de grado"
                        else:
                            mensaje = mensaje + ", con el archivo de su acta de grado"

                    if documentos.estadocedularepresentantesol == 3:
                        if mensaje == "":
                            mensaje = " se presentaron novedades con la revisión del archivo de la cédula del representante solidario"
                        else:
                            mensaje = mensaje + ", el archivo de la cédula del representante solidario"

                    if documentos.estadopapeletarepresentantesol == 3:
                        if mensaje == "":
                            mensaje = " se presentaron novedades con la revisión del archivo del certificado de votación del representante solidario"
                        else:
                            mensaje = mensaje + ", con el archivo del certificado de votación del representante solidario"

                    observaciones = documentos.observacion

                    recorrido = BecaSolicitudRecorrido(solicitud=beca.solicitud,
                                                       observacion=observaciones,
                                                       estado=12,
                                                       fecha=datetime.now().date())
                    recorrido.save(request)

                    send_html_mail(tituloemail,
                                   "emails/notificarrevisiondocumento.html",
                                   {'sistema': u'SGA - UNEMI',
                                    'mensaje': mensaje,
                                    'fecha': datetime.now().date(),
                                    'hora': datetime.now().time(),
                                    'observaciones': observaciones,
                                    'saludo': 'Estimada' if beca.solicitud.inscripcion.persona.sexo_id == 1 else 'Estimado',
                                    'estudiante': beca.solicitud.inscripcion.persona.nombre_completo_inverso(),
                                    'autoridad2': '',
                                    't': miinstitucion()
                                    },
                                   beca.solicitud.inscripcion.persona.lista_emails_envio(),
                                   [],
                                   cuenta=variable_valor('CUENTAS_CORREOS')[0]
                                   )
                else:
                    recorrido = BecaSolicitudRecorrido(solicitud=beca.solicitud,
                                                       observacion="DOCUMENTACIÓN VALIDADA",
                                                       estado=11,
                                                       fecha=datetime.now().date())
                    recorrido.save(request)
                    tituloemail = "Beca - Documentación Validada"

                    if beca.solicitud.becatipo.id == 23:
                        mensaje = "los documentos correspondientes a la cédula y certificado de votación del estudiante y del representante solidario han sido Validados"
                        observaciones = "Usted deberá acceder al Módulo de Beca Estudiantil, descargar el contrato de ayuda económica, imprimirlo, recoger las firmas correspondientes y subirlo al sistema en formato PDF."
                    else:
                        if beca.solicitud.becatipo.id == 16:
                            mensaje = "los documentos correspondientes al estudiante y representante solidario han sido Validados"
                        else:
                            mensaje = "los documentos correspondientes a la cédula y certificado de votación del estudiante y del representante solidario han sido Validados"

                        observaciones = "Usted deberá acceder al Módulo de Beca Estudiantil, descargar el contrato de beca (disponible a partir de las 08h00 del Martes 04 de Agosto del 2020), imprimirlo, recoger las firmas correspondientes y subirlo al sistema en formato PDF."

                    send_html_mail(tituloemail,
                                   "emails/notificarrevisiondocumento.html",
                                   {'sistema': u'SGA - UNEMI',
                                    'mensaje': mensaje,
                                    'fecha': datetime.now().date(),
                                    'hora': datetime.now().time(),
                                    'observaciones': observaciones,
                                    'saludo': 'Estimada' if beca.solicitud.inscripcion.persona.sexo_id == 1 else 'Estimado',
                                    'estudiante': beca.solicitud.inscripcion.persona.nombre_completo_inverso(),
                                    'autoridad2': '',
                                    't': miinstitucion()
                                    },
                                   beca.solicitud.inscripcion.persona.lista_emails_envio(),
                                   [],
                                   cuenta=variable_valor('CUENTAS_CORREOS')[0]
                                   )

                log(u'Revisó documentación cédula de la beca: %s  - %s' % (persona, beca), request, "edit")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'validarcontrato':
            try:
                beca = BecaAsignacion.objects.get(pk=int(request.POST['id']))
                beca.estadorevisioncontrato = int(request.POST['estadocontrato'])
                beca.observacion = request.POST['observacion'].strip().upper() if 'observacion' in request.POST else ''
                beca.save(request)

                if beca.estadorevisioncontrato == 3:
                    tituloemail = "Novedades con el Contrato de Beca"
                    mensaje = " se presentaron novedades con la revisión del archivo del contrato de beca"
                    observaciones = beca.observacion
                    recorrido = BecaSolicitudRecorrido(solicitud=beca.solicitud,
                                                       observacion=observaciones,
                                                       estado=20,
                                                       fecha=datetime.now().date())
                    recorrido.save(request)

                    send_html_mail(tituloemail,
                                   "emails/notificarrevisiondocumento.html",
                                   {'sistema': u'SGA - UNEMI',
                                    'mensaje': mensaje,
                                    'fecha': datetime.now().date(),
                                    'hora': datetime.now().time(),
                                    'observaciones': observaciones,
                                    'saludo': 'Estimada' if beca.solicitud.inscripcion.persona.sexo_id == 1 else 'Estimado',
                                    'estudiante': beca.solicitud.inscripcion.persona.nombre_completo_inverso(),
                                    'autoridad2': '',
                                    't': miinstitucion()
                                    },
                                   beca.solicitud.inscripcion.persona.lista_emails_envio(),
                                   [],
                                   cuenta=variable_valor('CUENTAS_CORREOS')[0]
                                   )
                else:
                    recorrido = BecaSolicitudRecorrido(solicitud=beca.solicitud,
                                                       observacion="CONTRATO DE BECA VALIDADO",
                                                       estado=19,
                                                       fecha=datetime.now().date())
                    recorrido.save(request)
                    tituloemail = "Beca - Contrato Validado"
                    mensaje = "el archivo correspondiente al contrato de la beca ha sido Validado"
                    observaciones = ""
                    send_html_mail(tituloemail,
                                   "emails/notificarrevisiondocumento.html",
                                   {'sistema': u'SGA - UNEMI',
                                    'mensaje': mensaje,
                                    'fecha': datetime.now().date(),
                                    'hora': datetime.now().time(),
                                    'observaciones': observaciones,
                                    'saludo': 'Estimada' if beca.solicitud.inscripcion.persona.sexo_id == 1 else 'Estimado',
                                    'estudiante': beca.solicitud.inscripcion.persona.nombre_completo_inverso(),
                                    'autoridad2': '',
                                    't': miinstitucion()
                                    },
                                   beca.solicitud.inscripcion.persona.lista_emails_envio(),
                                   [],
                                   cuenta=variable_valor('CUENTAS_CORREOS')[0]
                                   )

                log(u'Revisó documentación cédula de la beca: %s  - %s' % (persona, beca), request, "edit")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'validarcumplimiento':
            try:
                comprobante = BecaComprobanteVenta.objects.get(pk=int(request.POST['idcomprobante']))
                tipovalida = request.POST['tvalida']
                revisioncomprobante = comprobante.revision
                beca = revisioncomprobante.asignacion

                f = BecaComprobanteVentaValidaForm(request.POST)
                if f.is_valid():
                    if len(f.cleaned_data['rucproveedor'].strip()) != 13:
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"El número de Ruc del proveedor debe tener 13 dígitos."})

                    if f.cleaned_data['rucproveedor'].strip()[-3:] != '001':
                        return JsonResponse({"result": "bad", "mensaje": "El número de Ruc no es válido"})

                    if f.cleaned_data['total'] < 1:
                        return JsonResponse({"result": "bad", "mensaje": "El total del comprobante debe ser mayor a 0"})

                    comprobante.rucproveedor = f.cleaned_data['rucproveedor'].strip()
                    comprobante.total = f.cleaned_data['total']
                    if tipovalida == 'ADQ':
                        comprobante.estadorevisiondbu = int(f.cleaned_data['estado'])
                        comprobante.observaciondbu = f.cleaned_data['observacion'].strip().upper() if f.cleaned_data[
                            'observacion'] else ''
                    else:
                        comprobante.estadorevisionfin = int(f.cleaned_data['estado'])
                        comprobante.observacionfin = f.cleaned_data['observacion'].strip().upper() if f.cleaned_data[
                            'observacion'] else ''

                    comprobante.save(request)

                    if tipovalida == 'ADQ':
                        if comprobante.estadorevisiondbu == 3 or comprobante.estadorevisiondbu == 6:
                            comprobante.estado = 3
                            comprobante.save(request)
                            if comprobante.estadorevisiondbu == 3:
                                tituloemail = "Novedades con el Comprobante de Venta"
                                mensaje = " se presentaron novedades con la revisión del comprobante de venta"
                                observaciones = comprobante.observaciondbu
                                recorrido = BecaSolicitudRecorrido(solicitud=beca.solicitud,
                                                                   observacion=observaciones,
                                                                   estado=24,
                                                                   fecha=datetime.now().date())
                            else:
                                tituloemail = "TITULO X INCUMPLIMIENTO"
                                tituloemail = "Novedades con el Comprobante de Venta"
                                mensaje = " se presentaron novedades con la revisión del comprobante de venta"
                                observaciones = comprobante.observaciondbu
                                recorrido = BecaSolicitudRecorrido(solicitud=beca.solicitud,
                                                                   observacion=observaciones,
                                                                   estado=24,
                                                                   fecha=datetime.now().date())

                            recorrido.save(request)

                            send_html_mail(tituloemail,
                                           "emails/notificarrevisiondocumento.html",
                                           {'sistema': u'SGA - UNEMI',
                                            'mensaje': mensaje,
                                            'fecha': datetime.now().date(),
                                            'hora': datetime.now().time(),
                                            'observaciones': observaciones,
                                            'saludo': 'Estimada' if beca.solicitud.inscripcion.persona.sexo_id == 1 else 'Estimado',
                                            'estudiante': beca.solicitud.inscripcion.persona.nombre_completo_inverso(),
                                            'autoridad2': '',
                                            't': miinstitucion()
                                            },
                                           beca.solicitud.inscripcion.persona.lista_emails_envio(),
                                           [],
                                           cuenta=variable_valor('CUENTAS_CORREOS')[0]
                                           )
                        else:
                            comprobante.estado = 2
                            comprobante.save(request)

                            recorrido = BecaSolicitudRecorrido(solicitud=beca.solicitud,
                                                               observacion="COMPROBANTE VALIDADO POR BIENESTAR",
                                                               estado=23,
                                                               fecha=datetime.now().date())
                            recorrido.save(request)
                            tituloemail = "Beca - Comprobante de Venta Validado por Bienestar"
                            mensaje = "el archivo correspondiente al comprobante de venta ha sido Validado"
                            observaciones = ""
                            send_html_mail(tituloemail,
                                           "emails/notificarrevisiondocumento.html",
                                           {'sistema': u'SGA - UNEMI',
                                            'mensaje': mensaje,
                                            'fecha': datetime.now().date(),
                                            'hora': datetime.now().time(),
                                            'observaciones': observaciones,
                                            'saludo': 'Estimada' if beca.solicitud.inscripcion.persona.sexo_id == 1 else 'Estimado',
                                            'estudiante': beca.solicitud.inscripcion.persona.nombre_completo_inverso(),
                                            'autoridad2': '',
                                            't': miinstitucion()
                                            },
                                           beca.solicitud.inscripcion.persona.lista_emails_envio(),
                                           [],
                                           cuenta=variable_valor('CUENTAS_CORREOS')[0]
                                           )

                        comprobantes = revisioncomprobante.becacomprobanteventa_set.filter(status=True)
                        totalvalidado = comprobantes.filter(estadorevisiondbu=2).count()
                        totalrechazado = comprobantes.filter(estadorevisiondbu=3).count()
                        totalrechazadoio = comprobantes.filter(estadorevisiondbu=6).count()
                        totalcargado = comprobantes.filter(estado=1).count()

                        # if totalcargado == 0:
                        if totalrechazado == 0 and totalrechazadoio == 0 and totalcargado == 0:
                            revisioncomprobante.estado = 2
                            revisioncomprobante.estadorevisiondbu = 2
                            revisioncomprobante.estadorevisionfin = 1
                            revisioncomprobante.observaciondbu = ''
                            revisioncomprobante.observacionfin = ''
                        elif totalrechazado > 0 or totalrechazadoio > 0:
                            revisioncomprobante.estado = 3
                            revisioncomprobante.estadorevisiondbu = 3
                            observaciones = ",".join(
                                [c.observaciondbu for c in comprobantes.filter(estadorevisiondbu__in=[3, 6])])
                            revisioncomprobante.observaciondbu = observaciones

                        revisioncomprobante.save(request)
                    else:
                        if comprobante.estadorevisionfin == 3 or comprobante.estadorevisionfin == 6:
                            comprobante.estado = 6
                            comprobante.save(request)
                            if comprobante.estadorevisionfin == 3:
                                tituloemail = "Novedades con el Comprobante de Venta"
                                mensaje = " se presentaron novedades con la revisión del comprobante de venta"
                                observaciones = comprobante.observacionfin
                                recorrido = BecaSolicitudRecorrido(solicitud=beca.solicitud,
                                                                   observacion=observaciones,
                                                                   estado=27,
                                                                   fecha=datetime.now().date())
                            else:
                                tituloemail = "TITULO X INCUMPLIMIENTO"
                                tituloemail = "Novedades con el Comprobante de Venta"
                                mensaje = " se presentaron novedades con la revisión del comprobante de venta"
                                observaciones = comprobante.observacionfin
                                recorrido = BecaSolicitudRecorrido(solicitud=beca.solicitud,
                                                                   observacion=observaciones,
                                                                   estado=24,
                                                                   fecha=datetime.now().date())

                            recorrido.save(request)

                            send_html_mail(tituloemail,
                                           "emails/notificarrevisiondocumento.html",
                                           {'sistema': u'SGA - UNEMI',
                                            'mensaje': mensaje,
                                            'fecha': datetime.now().date(),
                                            'hora': datetime.now().time(),
                                            'observaciones': observaciones,
                                            'saludo': 'Estimada' if beca.solicitud.inscripcion.persona.sexo_id == 1 else 'Estimado',
                                            'estudiante': beca.solicitud.inscripcion.persona.nombre_completo_inverso(),
                                            'autoridad2': '',
                                            't': miinstitucion()
                                            },
                                           beca.solicitud.inscripcion.persona.lista_emails_envio(),
                                           [],
                                           cuenta=variable_valor('CUENTAS_CORREOS')[0]
                                           )
                        else:
                            comprobante.estado = 5
                            comprobante.save(request)

                            recorrido = BecaSolicitudRecorrido(solicitud=beca.solicitud,
                                                               observacion="COMPROBANTE VALIDADO POR FINANCIERO",
                                                               estado=26,
                                                               fecha=datetime.now().date())
                            recorrido.save(request)
                            tituloemail = "Beca - Comprobante de Venta Validado por Financiero"
                            mensaje = "el archivo correspondiente al comprobante de venta ha sido Validado"
                            observaciones = ""
                            send_html_mail(tituloemail,
                                           "emails/notificarrevisiondocumento.html",
                                           {'sistema': u'SGA - UNEMI',
                                            'mensaje': mensaje,
                                            'fecha': datetime.now().date(),
                                            'hora': datetime.now().time(),
                                            'observaciones': observaciones,
                                            'saludo': 'Estimada' if beca.solicitud.inscripcion.persona.sexo_id == 1 else 'Estimado',
                                            'estudiante': beca.solicitud.inscripcion.persona.nombre_completo_inverso(),
                                            'autoridad2': '',
                                            't': miinstitucion()
                                            },
                                           beca.solicitud.inscripcion.persona.lista_emails_envio(),
                                           [],
                                           cuenta=variable_valor('CUENTAS_CORREOS')[0]
                                           )

                        comprobantes = revisioncomprobante.becacomprobanteventa_set.filter(status=True)
                        totalvalidado = comprobantes.filter(estadorevisionfin=2).count()
                        totalrechazado = comprobantes.filter(estadorevisionfin=3).count()
                        totalrechazadoio = comprobantes.filter(estadorevisionfin=6).count()
                        totalcargado = comprobantes.filter(estadorevisionfin=1).count()

                        # if totalcargado == 0:
                        if totalrechazado == 0 and totalrechazadoio == 0 and totalcargado == 0:
                            revisioncomprobante.estado = 5
                            revisioncomprobante.estadorevisionfin = 2
                            revisioncomprobante.observacionfin = ''
                        elif totalrechazado > 0 or totalrechazadoio > 0:
                            revisioncomprobante.estado = 6
                            revisioncomprobante.estadorevisionfin = 3
                            observaciones = ",".join(
                                [c.observacionfin for c in comprobantes.filter(estadorevisionfin__in=[3, 6])])
                            revisioncomprobante.observacionfin = observaciones

                        revisioncomprobante.save(request)

                    log(u'Revisó documentación comprobante de venta de la beca: %s  - %s' % (persona, beca), request,
                        "edit")

                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'delcomprobante':
            try:
                comprobante = BecaComprobanteVenta.objects.get(pk=int(request.POST['idcomprobanteeli']))

                f = BecaComprobanteEliminarForm(request.POST)
                if f.is_valid():
                    comprobante.status = False
                    comprobante.observaciondbu = f.cleaned_data['observacioneli'].strip().upper()
                    comprobante.save(request)
                    log(u'%s eliminó comprobante de venta %s de la beca: %s' % (
                        persona, comprobante, comprobante.revision.asignacion), request, "del")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'representantesolidario':
            try:
                beca = BecaAsignacion.objects.get(pk=int(request.POST['id']))
                estudiante = beca.solicitud.inscripcion.persona
                representante = estudiante.personaextension_set.all()[0]

                f = RepresentanteSolidarioForm(request.POST)
                if f.is_valid():
                    if not f.cleaned_data['cedula'].isdigit():
                        return JsonResponse({"result": "bad", "mensaje": "El número de cédula debe ser numérico"})

                    resp = validarcedula(f.cleaned_data['cedula'])
                    if resp != 'Ok':
                        return JsonResponse({"result": "bad", "mensaje": resp})

                    if f.cleaned_data['cedula'] == estudiante.cedula:
                        return JsonResponse({"result": "bad",
                                             "mensaje": "El número de cédula del representante solidario debe ser distinto al del estudiante"})

                    representante.cedularepsolidario = f.cleaned_data['cedula']
                    representante.nombresrepsolidario = f.cleaned_data['nombre']
                    representante.apellido1repsolidario = f.cleaned_data['apellido1']
                    representante.apellido2repsolidario = f.cleaned_data['apellido2']
                    representante.save(request)

                    log(u'Actualizó datos de representante solidario: %s' % persona, request, "edit")
                    return JsonResponse({'result': 'ok'})
                else:
                    raise NameError('Error')

            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'subirarchivopendiente':
            try:
                solicitud = BecaSolicitud.objects.get(pk=int(request.POST['id']))

                if 'archivo1' in request.FILES:
                    arch = request.FILES['archivo1']
                    extension = arch._name.split('.')
                    tam = len(extension)
                    exte = extension[tam - 1]
                    if arch.size > 4194304:
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"Error, el tamaño del archivo es mayor a 4 Mb."})
                    if not exte.lower() == 'pdf':
                        return JsonResponse({"result": "bad", "mensaje": u"Solo se permiten archivos .pdf"})

                if 'archivo2' in request.FILES:
                    arch = request.FILES['archivo2']
                    extension = arch._name.split('.')
                    tam = len(extension)
                    exte = extension[tam - 1]
                    if arch.size > 4194304:
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"Error, el tamaño del archivo es mayor a 4 Mb."})
                    if not exte.lower() == 'pdf':
                        return JsonResponse({"result": "bad", "mensaje": u"Solo se permiten archivos .pdf"})

                detalle = solicitud.becadetallesolicitud_set.get(requisito_id=15)
                detalle.observacion = request.POST['observacion1'].strip().upper()
                newfile = request.FILES['archivo1']
                newfile._name = generar_nombre("requisito", newfile._name)
                detalle.archivo = newfile
                detalle.save(request)

                detalle = solicitud.becadetallesolicitud_set.get(requisito_id=16)
                detalle.observacion = request.POST['observacion2'].strip().upper()
                newfile = request.FILES['archivo2']
                newfile._name = generar_nombre("requisito", newfile._name)
                detalle.archivo = newfile
                detalle.save(request)

                log(u'Actualizó archivo de requisitos: %s  - %s' % (persona, solicitud), request, "edit")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'subirevidenciasentrega':
            try:
                beca = BecaAsignacion.objects.get(pk=int(request.POST['id']))

                if 'archivocontrato' in request.FILES:
                    arch = request.FILES['archivocontrato']
                    extension = arch._name.split('.')
                    tam = len(extension)
                    exte = extension[tam - 1]
                    if arch.size > 4194304:
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"Error, el tamaño del archivo del contrato es mayor a 4 Mb."})
                    if not exte.lower() == 'pdf':
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"Solo se permiten archivos .pdf en el contrato"})

                if 'archivoacta' in request.FILES:
                    arch = request.FILES['archivoacta']
                    extension = arch._name.split('.')
                    tam = len(extension)
                    exte = extension[tam - 1]
                    if arch.size > 4194304:
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"Error, el tamaño del archivo del acta es mayor a 4 Mb."})
                    if not exte.lower() == 'pdf':
                        return JsonResponse({"result": "bad", "mensaje": u"Solo se permiten archivos .pdf en el acta"})

                if 'archivofoto' in request.FILES:
                    arch = request.FILES['archivofoto']
                    extension = arch._name.split('.')
                    tam = len(extension)
                    exte = extension[tam - 1]
                    if arch.size > 4194304:
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"Error, el tamaño del archivo de la foto es mayor a 4 Mb."})
                    if not exte.lower() in ['jpg', 'jpeg', 'png']:
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"Solo se permiten archivos .jpg, .jpeg y .png en la foto"})

                if 'archivocontrato' in request.FILES:
                    newfile = request.FILES['archivocontrato']
                    newfile._name = generar_nombre("contratobeca", newfile._name)
                    beca.archivocontrato = newfile

                if 'archivoacta' in request.FILES:
                    newfile = request.FILES['archivoacta']
                    newfile._name = generar_nombre("actaentrega", newfile._name)
                    beca.archivoacta = newfile

                if 'archivofoto' in request.FILES:
                    newfile = request.FILES['archivofoto']
                    newfile._name = generar_nombre("fotoentrega", newfile._name)
                    beca.fotoentrega = newfile

                beca.save(request)

                log(u'Actualizó archivos de evidencias de beca: %s  - %s' % (persona, beca), request, "edit")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'asignarbeca2':
            try:
                solicitud = BecaSolicitud.objects.get(pk=int(request.POST['id']))
                beneficio = int(request.POST['beneficio'])

                if solicitud.becado_asignacion():
                    return JsonResponse(
                        {"result": "bad", "mensaje": u"Las asignación de la Beca ya sido realizada por otro usuario."})

                LIMITE_BECAS_PRESENCIAL_SEMIPRESENCIAL = 1150  # 1150
                LIMITE_BECAS_ENLINEA = 450  # 450

                becas = BecaAsignacion.objects.filter(solicitud__periodo=periodo, status=True,
                                                      solicitud__estado=2).exclude(aceptada=3)
                totalpresemi = becas.filter(solicitud__inscripcion__modalidad_id__in=[1, 2]).count()
                totalenlinea = becas.filter(solicitud__inscripcion__modalidad_id=3).count()

                if solicitud.inscripcion.modalidad.id in [1, 2]:
                    if totalpresemi == LIMITE_BECAS_PRESENCIAL_SEMIPRESENCIAL:
                        return JsonResponse({"result": "bad",
                                             "mensaje": u"El cupo para asignación de becas para TABLET + PLAN DE DATOS se agotó."})
                else:
                    if totalenlinea == LIMITE_BECAS_ENLINEA:
                        return JsonResponse({"result": "bad",
                                             "mensaje": u"El cupo para asignación de becas para PLAN DE DATOS se agotó."})

                f = BecaAsignacion2Form(request.POST)

                if f.is_valid():
                    beca = BecaAsignacion(solicitud=solicitud,
                                          montomensual=None,
                                          cantidadmeses=None,
                                          montobeneficio=None,
                                          fecha=datetime.now().date(),
                                          activo=True,
                                          grupopago=None,
                                          tipo=1,
                                          notificar=True,
                                          estadobeca=None,
                                          aceptada=1)
                    beca.save(request)

                    detalleuso = BecaDetalleUtilizacion(asignacion=beca,
                                                        utilizacion_id=8 if beneficio == 1 else 9,
                                                        personaaprueba=persona,
                                                        archivo=None,
                                                        fechaaprueba=datetime.now(),
                                                        fechaarchivo=None,
                                                        estado=2,
                                                        observacion='GENERADO AUTOMÁTICAMENTE')
                    detalleuso.save(request)

                    solicitud.becaasignada = 2
                    solicitud.save(request)

                    # Envio de e-mail de notificación al solicitante para que se pueda matricular
                    tipobeca = beca.solicitud.becatipo.nombre.upper()
                    tituloemail = "Asignación de Beca por " + tipobeca

                    send_html_mail(tituloemail,
                                   "emails/notificarestadosolicitudbeca.html",
                                   {'sistema': u'SGA - UNEMI',
                                    'fase': 'BECA',
                                    'tipobeca': tipobeca,
                                    'fecha': datetime.now().date(),
                                    'hora': datetime.now().time(),
                                    'saludo': 'Estimada' if solicitud.inscripcion.persona.sexo_id == 1 else 'Estimado',
                                    'estudiante': solicitud.inscripcion.persona.nombre_completo_inverso(),
                                    'autoridad2': '',
                                    't': miinstitucion()
                                    },
                                   solicitud.inscripcion.persona.lista_emails_envio(),
                                   [],
                                   cuenta=variable_valor('CUENTAS_CORREOS')[0]
                                   )

                    log(u'Asignó beca a la solicitud: %s  - %s' % (persona, solicitud), request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'updategrupopago':
            try:
                beca = BecaAsignacion.objects.get(pk=int(request.POST['bid']))
                if request.POST['vc']:
                    valor = int(request.POST['vc'])
                    if valor > 999:
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"No puede establecer un grupo pago menor a" + str(999 + 1)})
                else:
                    valor = None
                valoranterior = beca.grupopago
                beca.grupopago = valor
                beca.save(request)
                log(u'Actualizo grupo pago en beca: %s grupo pago anterior: %s grupo pago actual: %s' % (
                    beca, str(valoranterior if valoranterior else ""), str(valor)), request, "add")
                return JsonResponse({'result': 'ok', 'valor': beca.grupopago})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', "mensaje": u"Digite solo números o en blanco."})

        elif action == 'updatetipo':
            try:
                beca = BecaAsignacion.objects.get(pk=int(request.POST['bid']))
                valor = int(request.POST['tipo'])
                if int(request.POST['tipo']) == 0:
                    valor = None
                valoranterior = beca.tipo
                beca.tipo = valor
                beca.save(request)
                log(u'Actualizo tipo en beca: %s tipo anterior: %s tipo actual: %s' % (
                    beca, str(valoranterior if valoranterior else ""), str(valor)), request, "add")
                return JsonResponse({'result': 'ok'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', "mensaje": u"Digite solo números o en blanco."})

        elif action == 'informepago':
            mensaje = "Problemas al generar el informe de pago de becas."
            try:
                btipo = int(request.POST['btipo'])
                tipo = int(request.POST['tipo'])
                grupo = int(request.POST['grupo'])
                desde = convertir_fecha(request.POST['desde'])
                hasta = convertir_fecha(request.POST['hasta'])
                becas = BecaAsignacion.objects.filter(solicitud__periodo=periodo, status=True).order_by(
                    'solicitud__inscripcion__persona__apellido1', 'solicitud__inscripcion__persona__apellido2',
                    'solicitud__inscripcion__persona__nombres')
                # inicio = convertir_fecha(request.POST['fini'])
                # fin = convertir_fecha(request.POST['ffin'])
                if btipo != 0:
                    data['becatipo'] = becatipo = BecaTipo.objects.get(pk=btipo)
                    becas = becas.filter(solicitud__becatipo=becatipo)

                if tipo != 0:
                    becas = becas.filter(tipo=tipo)
                    data['tipo'] = TIPO_NUE_RENO[tipo - 1]
                if grupo != 0:
                    becas = becas.filter(grupopago=grupo)
                # data['becas'] = becas = becas[desde:hasta]
                data['becas'] = becas
                contador_letra = ['PRIMER MES', 'SEGUNDO MES', 'TERCERO MES', 'CUARTO MES', 'QUINTO MES', 'SEXTO MES',
                                  'SEPTIMO MES', 'OCTAVO MES', 'NOVENO MES']
                data['valormensual'] = null_to_numeric(becas.aggregate(valor=Sum('montomensual'))['valor'])
                data['valormensual2'] = valormesnual2 = []
                data['total'] = null_to_numeric(becas.aggregate(valor=Sum('montobeneficio'))['valor'])
                distributivos = DistributivoPersona.objects.filter(denominacionpuesto=547, status=True,
                                                                   estadopuesto_id=1)
                data['distributivo'] = distributivos[0] if distributivos.values('id').exists() else None
                cantidadmeses = becas.aggregate(cantidadmeses=Max("cantidadmeses"))['cantidadmeses']
                data['totalcantidadmeses'] = cantidadmeses
                data['lista_meses'] = contador_letra[:cantidadmeses]
                vmes = 0
                for mes in range(cantidadmeses):
                    for bec in becas:
                        vmes += bec.montomensual
                    valormesnual2.insert(mes, vmes)
                    vmes = 0

                return conviert_html_to_pdf('adm_becas/listadopagobeca.html', {'pagesize': 'A4', 'data': data})
            except Exception as ex:
                return HttpResponseRedirect("/adm_becas?info=%s" % mensaje)

        elif action == 'informepagotodo':
            mensaje = "Problemas al generar el informe de pago de becas."
            try:
                data['becas'] = BecaAsignacion.objects.filter(solicitud__periodo=periodo, status=True).order_by(
                    'solicitud__inscripcion__persona__apellido1', 'solicitud__inscripcion__persona__apellido2',
                    'solicitud__inscripcion__persona__nombres')
                distributivos = DistributivoPersona.objects.filter(denominacionpuesto=547, status=True,
                                                                   estadopuesto_id=1)
                data['distributivo'] = distributivos[0] if distributivos.values('id').exists() else None
                return conviert_html_to_pdf('adm_becas/listadopagobecatodos.html', {'pagesize': 'A4', 'data': data})
            except Exception as ex:
                return HttpResponseRedirect("/adm_becas?info=%s" % mensaje)

        elif action == 'generarinformeasistencia':
            mensaje = "Problemas al generar el informe de asistencias."
            try:
                # fi = convertir_fecha(request.POST['fi'])
                # ff = convertir_fecha(request.POST['ff'])
                # desde = convertir_fecha(request.POST['desde'])
                # hasta = convertir_fecha(request.POST['hasta'])
                # becas = BecaAsignacion.objects.filter(solicitud__periodo=periodo, status=True).order_by('solicitud__inscripcion__persona__apellido1','solicitud__inscripcion__persona__apellido2', 'solicitud__inscripcion__persona__apellido2','solicitud__inscripcion__persona__nombres')
                # tipobeca = int(request.POST['tipobeca'])
                # tipo = int(request.POST['tipo'])
                # grupo = int(request.POST['grupo'])
                # if tipobeca != 0:
                #     becatipo = BecaTipo.objects.get(pk=tipobeca)
                #     becas = becas.filter(solicitud__becatipo=becatipo)
                # if tipo != 0:
                #     becas = becas.filter(tipo=tipo)
                # if grupo != 0:
                #     becas = becas.filter(grupopago=grupo)
                # #becas=becas[int(desde):int(hasta)]
                # distributivos = DistributivoPersona.objects.filter(denominacionpuesto=547, status=True, estadopuesto_id=1)
                # data['distributivo'] = distributivos[0] if distributivos.values('id').exists() else None
                # data['periodo'] = periodo
                # # data['fechainicio'] = fi
                # # data['fechafin'] = ff
                # listadobecas = []
                # contador = 0
                # for beca in becas:
                #     contador+=1
                #     listadobecas.append([contador.__str__(), beca.solicitud.inscripcion.persona.cedula, beca.solicitud.inscripcion.persona.nombre_completo_inverso(),beca.solicitud.inscripcion.mi_nivel(),beca.solicitud.inscripcion.carrera, beca.solicitud.becatipo, beca.solicitud.get_estado_display(), beca.total_promedio_asistencia(periodo.inicio,periodo.fin)])
                # data['becas'] = listadobecas

                becas = BecaAsignacion.objects.filter(solicitud__periodo=periodo, status=True).order_by(
                    'solicitud__inscripcion__persona__apellido1', 'solicitud__inscripcion__persona__apellido2',
                    'solicitud__inscripcion__persona__apellido2', 'solicitud__inscripcion__persona__nombres')
                for beca in becas:
                    if not BecaTablaAsistencia.objects.filter(cedula=beca.solicitud.inscripcion.persona.cedula,
                                                              status=True).exists():
                        becario = BecaTablaAsistencia(periodo=periodo, cedula=beca.solicitud.inscripcion.persona.cedula,
                                                      apellidos_nombres=beca.solicitud.inscripcion.persona.nombre_completo_inverso(),
                                                      nivel=beca.solicitud.inscripcion.mi_nivel(),
                                                      carrera=beca.solicitud.inscripcion.carrera,
                                                      tipobeca=beca.solicitud.becatipo,
                                                      tipo=beca.tipo,
                                                      estadobeca=beca.solicitud.get_estado_display(),
                                                      grupopago=beca.grupopago,
                                                      asistencia=beca.total_promedio_asistencia(periodo.inicio,
                                                                                                periodo.fin))

                        becario.save(request)
                        mensaje = "Generación completada."
                    else:
                        mensaje = "Registros Existentes"

                return HttpResponseRedirect("/adm_becas?info=%s" % mensaje)
            except Exception as ex:
                return HttpResponseRedirect("/adm_becas?info=%s" % mensaje)

        elif action == 'informeasistencia':
            mensaje = "Problemas al generar el informe de asistencias."
            try:
                fi = convertir_fecha(request.POST['fi'])
                ff = convertir_fecha(request.POST['ff'])
                becas = BecaTablaAsistencia.objects.filter(periodo=periodo).order_by('apellidos_nombres')
                tipobeca = int(request.POST['tipobeca'])
                tipo = int(request.POST['tipo'])
                grupo = int(request.POST['grupo'])
                if tipobeca != 0:
                    becatipo = BecaTipo.objects.get(pk=tipobeca)
                    becas = becas.filter(tipobeca=becatipo)
                if tipo != 0:
                    becas = becas.filter(tipo=tipo)
                if grupo != 0:
                    becas = becas.filter(grupopago=grupo)
                distributivos = DistributivoPersona.objects.filter(denominacionpuesto=547, status=True,
                                                                   estadopuesto_id=1)
                data['distributivo'] = distributivos[0] if distributivos.values('id').exists() else None
                data['periodo'] = periodo
                data['fechainicio'] = fi
                data['fechafin'] = ff
                listadobecas = []
                contador = 0
                for beca in becas:
                    contador += 1
                    listadobecas.append(
                        [contador.__str__(), beca.cedula, beca.apellidos_nombres, beca.nivel, beca.carrera,
                         beca.tipobeca, beca.estadobeca, beca.asistencia])
                data['becas'] = listadobecas
                return conviert_html_to_pdf('adm_becas/informe_asistencia_becado.html',
                                            {'pagesize': 'A4', 'data': data})
            except Exception as ex:
                return HttpResponseRedirect("/adm_becas?info=%s" % mensaje)

        elif action == 'informepagoporpartes':
            mensaje = "Problemas al generar el informe de pago de becas."
            try:
                btipo = int(request.POST['btipo'])
                tipo = int(request.POST['tipo'])
                grupo = int(request.POST['grupo'])
                min_mes = int(request.POST['min_mes'])
                max_mes = int(request.POST['max_mes'])
                becas = BecaAsignacion.objects.filter(solicitud__periodo=periodo, status=True)
                # inicio = convertir_fecha(request.POST['fini'])
                # fin = convertir_fecha(request.POST['ffin'])
                if btipo != 0:
                    data['becatipo'] = becatipo = BecaTipo.objects.get(pk=btipo)
                    becas = becas.filter(solicitud__becatipo=becatipo)

                if tipo != 0:
                    becas = becas.filter(tipo=tipo)
                    data['tipo'] = TIPO_NUE_RENO[tipo - 1]
                if grupo != 0:
                    becas = becas.filter(grupopago=grupo)
                data['becas'] = becas.order_by('solicitud__inscripcion__persona__apellido1',
                                               'solicitud__inscripcion__persona__apellido2',
                                               'solicitud__inscripcion__persona__nombres')
                data['valormensual'] = valormensual = null_to_numeric(
                    becas.aggregate(valor=Sum('montomensual'))['valor'])
                distributivos = DistributivoPersona.objects.filter(denominacionpuesto=547, status=True,
                                                                   estadopuesto_id=1)
                data['distributivo'] = distributivos[0] if distributivos.values('id').exists() else None
                contador_letra = ['PRIMER MES', 'SEGUNDO MES', 'TERCERO MES', 'CUARTO MES', 'QUINTO MES', 'SEXTO MES',
                                  'SEPTIMO MES', 'OCTAVO MES', 'NOVENO MES']
                data['lista_meses'] = lista_meses = contador_letra[min_mes - 1:max_mes]
                data['totalcantidadmeses'] = cantmeses = lista_meses.__len__()
                data['total'] = null_to_decimal(valormensual * cantmeses, 2)
                return conviert_html_to_pdf('adm_becas/listadopagobecaporpartes.html', {'pagesize': 'A4', 'data': data})
            except Exception as ex:
                return HttpResponseRedirect("/adm_becas?info=%s" % mensaje)

        elif action == 'notificar':
            try:
                asunto = u"NOTIFICACION DE BECA"
                id = int(request.POST['id'])
                becado = BecaAsignacion.objects.get(id=id, solicitud__periodo=periodo, status=True)
                send_html_mail(asunto, "emails/notificarbecado.html",
                               {'sistema': request.session['nombresistema'],
                                'alumno': becado.solicitud.inscripcion.persona.nombre_completo_inverso()},
                               becado.solicitud.inscripcion.persona.lista_emails_envio(), [],
                               cuenta=CUENTAS_CORREOS[0][1])
                return JsonResponse({'result': 'ok'})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al notoficar."})

        elif action == 'notificartodo':
            try:
                asunto = u"NOTIFICACION DE BECA"
                mensaje = (request.POST['notificacion'])
                # becados = BecaAsignacion.objects.filter(solicitud__periodo=periodo, status=True)
                # becaperiodo = BecaPeriodo.objects.get(status= True, periodo=periodo)
                lista = []
                lista.append("jvaldezj@unemi.edu.ec")
                lista.append("jplacesc@unemi.edu.ec")
                # for becado in becados:
                #     send_html_mail(asunto, "emails/notificarbecado.html",
                #                {'sistema': request.session['nombresistema'],
                #                 'alumno': becado.solicitud.inscripcion.persona.nombre_completo_inverso(),
                #                 'fechafinbeca': becaperiodo.fechafinnotificacion},
                #                 lista, [],
                #                cuenta=CUENTAS_CORREOS[0][1])
                #     becado.notificar = True
                becado = BecaAsignacion.objects.get(solicitud__periodo=periodo, status=True, id=6123)
                send_html_mail(asunto, "emails/notificarbecado.html",
                               {'sistema': request.session['nombresistema'],
                                'alumno': becado.solicitud.inscripcion.persona.nombre_completo_inverso(),
                                'mensaje': mensaje},
                               lista, [],
                               cuenta=CUENTAS_CORREOS[8][1])

                return JsonResponse({'result': 'ok'})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al notificar." % ex})

        elif action == 'addtipobeca':
            try:
                f = BecaTipoForm(request.POST)
                if f.is_valid():
                    becatipo = BecaTipo(nombre=f.cleaned_data['nombre'],
                                        minimo_asistencia=f.cleaned_data['minimo_asistencia'],
                                        minimo_promedio=f.cleaned_data['minimo_promedio'],
                                        numemejores=f.cleaned_data['numemejores'],
                                        vigente=f.cleaned_data['vigente'],
                                        nombrecorto=f.cleaned_data['nombrecorto'],
                                        nombrecaces=f.cleaned_data['nombrecaces'],
                                        )
                    becatipo.save(request)
                    log(u'Agrego un Tipo Beca: %s' % becatipo, request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'edittipobeca':
            try:
                tipobeca = BecaTipo.objects.get(pk=int(request.POST['id']), status=True)
                f = BecaTipoForm(request.POST)
                if f.is_valid():
                    tipobeca.nombre = f.cleaned_data['nombre']
                    tipobeca.nombrecorto = f.cleaned_data['nombrecorto']
                    tipobeca.minimo_asistencia = f.cleaned_data['minimo_asistencia']
                    tipobeca.minimo_promedio = f.cleaned_data['minimo_promedio']
                    tipobeca.numemejores = f.cleaned_data['numemejores']
                    tipobeca.vigente = f.cleaned_data['vigente']
                    if f.cleaned_data['nombrecaces']:
                        tipobeca.nombrecaces = f.cleaned_data['nombrecaces']
                    tipobeca.save(request)
                    log(u'Edita un Tipo Beca: %s' % tipobeca, request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'deletipobeca':
            try:
                if BecaTipo.objects.get(pk=int(request.POST['id']), status=True):
                    tipobeca = BecaTipo.objects.get(pk=int(request.POST['id']), status=True)
                    tipobeca.delete()
                    log(u'Elimino un Tipo Beca: %s' % tipobeca, request, "del")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'addutilizacion':
            try:
                f = BecaUtilizacionForm(request.POST)
                if f.is_valid():
                    becautilizacion = BecaUtilizacion(nombre=f.cleaned_data['nombre'],
                                                      vigente=f.cleaned_data['vigente'])
                    becautilizacion.save(request)
                    log(u'Agrego una Utilicación de Beca: %s' % becautilizacion, request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'editutilizacion':
            try:
                utilizacion = BecaUtilizacion.objects.get(pk=int(request.POST['id']), status=True)
                f = BecaUtilizacionForm(request.POST)
                if f.is_valid():
                    utilizacion.nombre = f.cleaned_data['nombre']
                    utilizacion.vigente = f.cleaned_data['vigente']
                    utilizacion.save(request)
                    log(u'Edita una Utilicación de Beca: %s' % utilizacion, request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'deleutilizacion':
            try:
                if BecaUtilizacion.objects.get(pk=int(request.POST['id']), status=True):
                    utilizacion = BecaUtilizacion.objects.get(pk=int(request.POST['id']), status=True)
                    utilizacion.delete()
                    log(u'Elimino una Utilicación de Beca: %s' % utilizacion, request, "del")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'addrequisito':
            try:
                f = BecaRequisitosForm(request.POST)
                if f.is_valid():
                    requisito = BecaRequisitos(nombre=f.cleaned_data['nombre'],
                                               vigente=f.cleaned_data['vigente'],
                                               matricula=f.cleaned_data['matricula'],
                                               regular=f.cleaned_data['regular'],
                                               residencia=f.cleaned_data['residencia'],
                                               reprobado=f.cleaned_data['reprobado'],
                                               nodeudar=f.cleaned_data['nodeudar']
                                               )
                    requisito.save(request)
                    log(u'Agrego un Requisito de Beca: %s' % requisito, request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'editrequisito':
            try:
                requisito = BecaRequisitos.objects.get(pk=int(request.POST['id']), status=True, vigente='True')
                f = BecaRequisitosForm(request.POST)
                if f.is_valid():
                    requisito.nombre = f.cleaned_data['nombre']
                    requisito.vigente = f.cleaned_data['vigente']
                    requisito.matricula = f.cleaned_data['matricula']
                    requisito.regular = f.cleaned_data['regular']
                    requisito.residencia = f.cleaned_data['residencia']
                    requisito.reprobado = f.cleaned_data['reprobado']
                    requisito.nodeudar = f.cleaned_data['nodeudar']
                    requisito.save(request)
                    log(u'Edita un Requisito de Beca: %s' % requisito, request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'delerequisito':
            try:
                if BecaRequisitos.objects.get(pk=int(request.POST['id']), status=True, vigente='True'):
                    requisito = BecaRequisitos.objects.get(pk=int(request.POST['id']), status=True, vigente='True')
                    requisito.delete()
                    log(u'Elimino un Requisito de Beca: %s' % requisito, request, "del")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'rechazarsolicitud':
            try:
                becasolicitud = BecaSolicitud.objects.get(pk=request.POST['id'], periodo=periodo)
                becasolicitud.estado = 3
                becasolicitud.save(request)
                for x in becasolicitud.becadetallesolicitud_set.all():
                    x.estado = 3
                    x.save(request)
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos.%s " % ex})

        elif action == 'addbecaperiodo':
            try:
                f = BecaPeriodoForm(request.POST)
                if not f.is_valid():
                    raise NameError('Formulario incompleto')
                if BecaPeriodo.objects.filter(periodo=f.cleaned_data['periodo']).exists():
                    raise NameError('Periodo académico existente.')
                becaperiodo = BecaPeriodo(periodo=f.cleaned_data['periodo'],
                                          fechainiciosolicitud=convertir_fecha_hora_invertida(
                                              str(f.cleaned_data['fechainiciosolicitud']) + ' 00:00:00'),
                                          fechafinsolicitud=convertir_fecha_hora_invertida(
                                              str(f.cleaned_data['fechafinsolicitud']) + ' 23:59:59'),
                                          fechainiciovalidaciondocumento=convertir_fecha_hora_invertida(
                                              str(f.cleaned_data['fechainiciovalidaciondocumento']) + ' 00:00:00'),
                                          fechafinvalidaciondocumento=convertir_fecha_hora_invertida(
                                              str(f.cleaned_data['fechafinvalidaciondocumento']) + ' 23:59:59'),
                                          fechainicioimprimircontrato=convertir_fecha_hora_invertida(
                                              str(f.cleaned_data['fechainicioimprimircontrato']) + ' ' + str(
                                                  f.cleaned_data['horainicioimprimircontrato'])),
                                          fechafinimprimircontrato=convertir_fecha_hora_invertida(
                                              str(f.cleaned_data['fechafinimprimircontrato']) + ' ' + str(
                                                  f.cleaned_data['horafinimprimircontrato'])),
                                          fechainiciovalidacioncontrato=convertir_fecha_hora_invertida(
                                              str(f.cleaned_data['fechainiciovalidacioncontrato']) + ' 00:00:00'),
                                          fechafinvalidacioncontrato=convertir_fecha_hora_invertida(
                                              str(f.cleaned_data['fechafinvalidacioncontrato']) + ' 23:59:59'),
                                          fechainicioactualizarcertificariobancario=convertir_fecha_hora_invertida(str(
                                              f.cleaned_data[
                                                  'fechainicioactualizarcertificariobancario']) + ' 00:00:00'),
                                          fechafinactualizarcertificariobancario=convertir_fecha_hora_invertida(str(
                                              f.cleaned_data['fechafinactualizarcertificariobancario']) + ' 23:59:59'),
                                          obligadosubircomprobante=f.cleaned_data['obligadosubircomprobante'],
                                          vigente=f.cleaned_data['vigente'],
                                          limitebecados=f.cleaned_data['limitebecados'],
                                          )
                periodo.aplicabeca = True
                periodo.save(request)
                becaperiodo.save(request)
                for data in f.cleaned_data['nivelesmalla']:
                    becaperiodo.nivelesmalla.add(data)
                log(u'Nuevo periodo beca: %s' % becaperiodo, request, "add")
                return JsonResponse({'result': 'ok'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', 'mensaje': u'Error al guardar los datos. </br> %s' % ex})

        elif action == 'editbecaperiodo':
            try:
                f = BecaPeriodoForm(request.POST)
                becaperiodo = BecaPeriodo.objects.get(pk=request.POST['id'])
                if not f.is_valid():
                    raise NameError('Error en el formulario')
                #becaperiodo.periodo = f.cleaned_data['periodo']
                becaperiodo.fechainiciosolicitud = convertir_fecha_hora_invertida(
                    str(f.cleaned_data['fechainiciosolicitud']) + ' 00:00:00')
                becaperiodo.fechafinsolicitud = convertir_fecha_hora_invertida(
                    str(f.cleaned_data['fechafinsolicitud']) + ' 23:59:59')
                becaperiodo.fechainiciovalidaciondocumento = convertir_fecha_hora_invertida(
                    str(f.cleaned_data['fechainiciovalidaciondocumento']) + ' 00:00:00')
                becaperiodo.fechafinvalidaciondocumento = convertir_fecha_hora_invertida(
                    str(f.cleaned_data['fechafinvalidaciondocumento']) + ' 23:59:59')
                becaperiodo.fechainicioimprimircontrato = convertir_fecha_hora_invertida(
                    str(f.cleaned_data['fechainicioimprimircontrato']) + ' ' + str(
                        f.cleaned_data['horainicioimprimircontrato']))
                becaperiodo.fechafinimprimircontrato = convertir_fecha_hora_invertida(
                    str(f.cleaned_data['fechafinimprimircontrato']) + ' ' + str(
                        f.cleaned_data['horafinimprimircontrato']))
                becaperiodo.fechainiciovalidacioncontrato = convertir_fecha_hora_invertida(
                    str(f.cleaned_data['fechainiciovalidacioncontrato']) + ' 00:00:00')
                becaperiodo.fechafinvalidacioncontrato = convertir_fecha_hora_invertida(
                    str(f.cleaned_data['fechafinvalidacioncontrato']) + ' 23:59:59')
                becaperiodo.fechainicioactualizarcertificariobancario = convertir_fecha_hora_invertida(
                    str(f.cleaned_data['fechainicioactualizarcertificariobancario']) + ' 00:00:00')
                becaperiodo.fechafinactualizarcertificariobancario = convertir_fecha_hora_invertida(
                    str(f.cleaned_data['fechafinactualizarcertificariobancario']) + ' 23:59:59')
                becaperiodo.obligadosubircomprobante = f.cleaned_data['obligadosubircomprobante']
                becaperiodo.vigente = f.cleaned_data['vigente']
                becaperiodo.limitebecados = f.cleaned_data['limitebecados']
                for data in f.cleaned_data['nivelesmalla']:
                    becaperiodo.nivelesmalla.add(data)
                becaperiodo.save(request)
                log(u'Modificó periodo beca: %s' % becaperiodo, request, "edit")
                return JsonResponse({'result': 'ok'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', 'mensaje': u'Error al guardar los datos'})

        elif action == 'imprimir_promedioobs':
            try:
                font_style = XFStyle()
                font_style.font.bold = True
                font_style2 = XFStyle()
                font_style2.font.bold = False
                wb = Workbook(encoding='utf-8')
                ws = wb.add_sheet('alumnos_promedio_2')
                response = HttpResponse(content_type="application/ms-excel")
                response['Content-Disposition'] = 'attachment; filename=alumnos_2' + random.randint(1,
                                                                                                    10000).__str__() + '.xls'
                columns = [
                    (u"N.", 1500),
                    (u"ID_PREINSCRIPCION.", 1500),
                    (u"APELLIDOS Y NOMBRES", 12000),
                    (u"CARRERA", 3000),
                    (u"CEDULA", 3000),
                    (u"FECHA NACIMIENTO", 3000),
                    (u"PROMEDIO", 2000),
                    (u"ASISTENCIA", 2000),
                    (u"SESION", 3000),
                    (u"NIVEL", 3000),
                    (u"PARALELO", 3000),
                    (u"OBERVACIÓN1", 3000),
                    (u"OBERVACIÓN2", 3000),
                    (u"DIRECCION", 8000),
                    (u"TELEFONO", 4000),
                    (u"TELEFONO CONVENCIONAL", 4000),
                    (u"EMAIL", 5000),
                    (u"EMAIL INSTITUCIONAL", 5000),
                    (u"SEXO", 5000)
                ]
                row_num = 0
                for col_num in range(len(columns)):
                    ws.write(row_num, col_num, columns[col_num][0], font_style)
                    ws.col(col_num).width = columns[col_num][1]
                date_format = xlwt.XFStyle()
                date_format.num_format_str = 'yyyy/mm/dd'
                inscripcionesactuales = Matricula.objects.values_list('inscripcion__id', flat=True).filter(
                    status=True, nivel__periodo__id=periodo.id,
                    estado_matricula__in=[2, 3],
                    retiradomatricula=False,
                    matriculagruposocioeconomico__tipomatricula=1).exclude(inscripcion__persona__rubro__cancelado=False,
                                                                           inscripcion__persona__rubro__status=True,
                                                                           inscripcion__persona__rubro__fecha__lte='2017-03-30').distinct().order_by(
                    "inscripcion__persona")
                nuevolistado = Matricula.objects.filter(status=True,
                                                        nivel__periodo__id=anterior.id,
                                                        estado_matricula__in=[2, 3],
                                                        retiradomatricula=False,
                                                        matriculagruposocioeconomico__tipomatricula=1,
                                                        inscripcion__id__in=inscripcionesactuales).exclude(
                    inscripcion__persona__rubro__cancelado=False,
                    inscripcion__persona__rubro__status=True,
                    inscripcion__persona__rubro__fecha__lte='2017-03-30').distinct().order_by("inscripcion__persona")
                row_num = 1
                i = 0
                for x in nuevolistado:
                    preinscripcion = PreInscripcionBeca.objects.filter(inscripcion=x.inscripcion, periodo=periodo).first()
                    observacion1 = None
                    observacion2 = None
                    materias = MateriaAsignada.objects.filter(status=True,
                                                              matricula__inscripcion__id=x.inscripcion.id,
                                                              matricula__nivel__periodo__id=anterior.id,
                                                              materiaasignadaretiro__isnull=True)
                    verifica = 0
                    suma = 0
                    sumasis = 0
                    promedio = 0
                    total = materias.count()
                    if total < 4:
                        observacion1 = "< 4"
                    for m in materias:
                        suma += m.notafinal
                        sumasis += m.asistenciafinal
                        if m.estado.id != 1:
                            verifica = 1
                            break
                    if suma > 0 and sumasis > 0:
                        promedio = round(suma / total, 2)
                        asistencia = round(sumasis / total, 2)
                    if verifica == 1 or promedio < 85:
                        observacion2 = "NO CUMPLE"
                    campo1 = preinscripcion.id
                    campo2 = x.inscripcion.persona.nombre_completo_inverso()
                    if (x.inscripcion.carrera.mencion):
                        campo3 = x.inscripcion.carrera.nombre + ' CON MENCION EN  ' + x.inscripcion.carrera.mencion
                    else:
                        campo3 = x.inscripcion.carrera.nombre
                    campo4 = x.inscripcion.persona.cedula
                    campo5 = promedio
                    campo6 = asistencia
                    campo7 = x.inscripcion.persona.nacimiento
                    campo8 = x.inscripcion.sesion.nombre
                    campo9 = x.nivelmalla.nombre
                    if (x.paralelo):
                        campo10 = x.paralelo.nombre
                    else:
                        campo10 = 'Ninguno'
                    campo11 = observacion1
                    campo12 = observacion2
                    campo13 = x.inscripcion.persona.direccion_completa()
                    campo14 = x.inscripcion.persona.telefono
                    campo15 = x.inscripcion.persona.telefono_conv
                    campo16 = x.inscripcion.persona.email
                    campo17 = x.inscripcion.persona.emailinst
                    campo18 = x.inscripcion.persona.sexo.nombre
                    i += 1
                    ws.write(row_num, 0, i, font_style2)
                    ws.write(row_num, 1, campo1, font_style2)
                    ws.write(row_num, 2, campo2, font_style2)
                    ws.write(row_num, 3, campo3, font_style2)
                    ws.write(row_num, 4, campo4, font_style2)
                    ws.write(row_num, 5, campo5, font_style2)
                    ws.write(row_num, 6, campo6, font_style2)
                    ws.write(row_num, 7, campo7, date_format)
                    ws.write(row_num, 8, campo8, font_style2)
                    ws.write(row_num, 9, campo9, font_style2)
                    ws.write(row_num, 10, campo10, font_style2)
                    ws.write(row_num, 11, campo11, font_style2)
                    ws.write(row_num, 12, campo12, font_style2)
                    ws.write(row_num, 13, campo13, font_style2)
                    ws.write(row_num, 14, campo14, font_style2)
                    ws.write(row_num, 15, campo15, font_style2)
                    ws.write(row_num, 16, campo16, font_style2)
                    ws.write(row_num, 17, campo17, font_style2)
                    ws.write(row_num, 18, campo18, font_style2)
                    row_num += 1
                wb.save(response)
                return response
            except Exception as ex:
                err = 'Error on line {}'.format(sys.exc_info()[-1].tb_lineno)
                msg = ex.__str__()
                msg = f'{msg} {err}'
                return JsonResponse({"result": "bad", "mensaje": u"Error: %s"%msg})

        elif action == 'imprimir_quintil':
            try:
                becatipo = BecaTipo.objects.get(pk=18)
                configuracion = becatipo.becatipoconfiguracion_set.filter(becaperiodo__periodo=periodo).first()

                action_extra = request.POST.get('a')
                cantidad_total_requisitos = 0
                nombre_archivo = 'alumnos_quintil'
                titulo_archivo = 'LISTADO DE ESTUDIANTE PRESELECCIONADO POR SITUACION ECONOMICA VULNERABLE'
                if configuracion:
                    if action_extra:
                        cantidad_total_requisitos = configuracion.requisitosbecas.count() if configuracion.requisitosbecas else 0
                        if action_extra == 'requisitos_completos':
                            nombre_archivo = 'alumnos_quintil_OCAS'
                            titulo_archivo = 'LISTADO DE ESTUDIANTES SELECCIONADOS POR SITUACION ECONOMICA VULNERABLE'
                        elif action_extra == 'requisitos_incompletos':
                            nombre_archivo = 'alumnos_quintil_rechazados'
                            titulo_archivo = 'LISTADO DE ESTUDIANTES NO SELECCIONADOS POR SITUACION ECONOMICA VULNERABLE'

                font_style = XFStyle()
                font_style.font.bold = True
                font_style2 = XFStyle()
                font_style2.font.bold = False
                wb = Workbook(encoding='utf-8')
                ws = wb.add_sheet('alumnos_quintil')
                # periodoactual = Periodo.objects.get(id=int(request.POST['periodoactual']))
                # periodovalida = Periodo.objects.get(id=int(request.POST['periodovalida']))
                periodoactual = periodosesion
                periodovalida = anterior
                # modalidad = int(request.POST['modalidad'])
                response = HttpResponse(content_type="application/ms-excel")
                response['Content-Disposition'] = f'attachment; filename={nombre_archivo}' + random.randint(1,
                                                                                                          10000).__str__() + '.xls'
                columns = [
                    (u"N.", 1500),
                    (u"ID_PREINSCRIPCION.", 1500),
                    (u"ORDEN", 1500),
                    (u"APELLIDOS Y NOMBRES", 12000),
                    (u"CARRERA", 3000),
                    (u"MODALIDAD", 3000),
                    (u"CEDULA", 3000),
                    (u"FECHA NACIMIENTO", 3000),
                    (u"DIRECCION", 3000),
                    (u"PROMEDIO GENERAL", 3000),
                    (u"PROMEDIO VERIFICADOR", 3000),
                    (u"ASISTENCIA", 3000),
                    (u"CODIGO", 3000),
                    (u"NOMBRE", 3000),
                    (u"SESION", 3000),
                    (u"NIVEL", 3000),
                    (u"PARALELO", 3000),
                    (u"DIRECCION COMPLETA", 3000),
                    (u"TELEFONO", 4000),
                    (u"TELEFONO CONVENCIONAL", 4000),
                    (u"EMAIL", 5000),
                    (u"EMAIL INSTITUCIONAL", 5000),
                    (u"SEXO", 5000),
                    (u"PAIS", 5000),
                    (u"PROVINCIA", 5000),
                    (u"CANTON", 5000),
                    (u"DIRECCION 1", 5000),
                    (u"DIRECCION 2", 5000),
                    (u"SECTOR", 5000),
                    (u"PERIODO ACTUAL", 8000),
                    (u"PERIDOO VALIDA", 8000),
                    (u"TIPO/ESTADO", 2500),
                ]
                requisitos = []
                if configuracion:
                    #Requisitos de tipos de becas
                    requisitos = configuracion.requisitosbecas.filter(visible=True, status=True)
                    for detallerequisitobeca in requisitos:
                        columns.append((f'({detallerequisitobeca.pk}) {detallerequisitobeca.requisitobeca.__str__()}', 10000))
                        if not action_extra:
                            columns.append((f'({detallerequisitobeca.pk}) OBSERVACIÓN', 5000))

                style_title = xlwt.easyxf(
                    'font: height 350, name Arial, colour_index black, bold on, italic on; align: wrap on, vert centre, horiz center;')
                style_title_2 = xlwt.easyxf(
                    'font: height 250, name Arial, colour_index black, bold on, italic on; align: wrap on, vert centre, horiz center;')
                ws.write_merge(0, 0, 0, len(columns), 'UNIVERSIDAD ESTATAL DE MILAGRO', style_title)
                ws.write_merge(1, 1, 0, len(columns), titulo_archivo, style_title_2)
                row_num = 3
                for col_num in range(len(columns)):
                    ws.write(row_num, col_num, columns[col_num][0], font_style)
                    ws.col(col_num).width = columns[col_num][1]
                date_format = xlwt.XFStyle()
                date_format.num_format_str = 'yyyy/mm/dd'
                matriculados = None
                if periodosesion.versionbeca is None or periodosesion.versionbeca == 1:
                    listaquintil = MatriculaGrupoSocioEconomico.objects.values_list('matricula__inscripcion__id',
                                                                                    flat=True).filter(
                        Q(gruposocioeconomico__codigo='D') | Q(gruposocioeconomico__codigo='C-'),
                        matricula__estado_matricula__in=[2, 3], matricula__status=True,
                        matricula__retiradomatricula=False,
                        matricula__nivel__periodo__id=periodo.id,
                        matricula__matriculagruposocioeconomico__tipomatricula=1,
                        matricula__inscripcion__carrera__coordinacion__excluir=False
                    ).exclude(matricula__inscripcion__carrera__coordinacion__id__in=[7, 9]).distinct().order_by(
                        "puntajetotal")
                    matriculados = MatriculaGrupoSocioEconomico.objects.filter(
                        Q(gruposocioeconomico__codigo='D') | Q(gruposocioeconomico__codigo='C-'),
                        matricula__estado_matricula__in=[2, 3], matricula__status=True,
                        matricula__retiradomatricula=False,
                        matricula__nivel__periodo__id=anterior.id,
                        matricula__matriculagruposocioeconomico__tipomatricula=1,
                        matricula__inscripcion__carrera__coordinacion__excluir=False,
                        matricula__inscripcion__id__in=listaquintil).exclude(
                        matricula__inscripcion__carrera__coordinacion__id__in=[7, 9]).distinct().order_by(
                        "puntajetotal")
                elif periodosesion.versionbeca == 2:
                    inscripciones = PreInscripcionBeca.objects.filter(periodo=periodo, becatipo_id=18)

                    if action_extra == 'requisitos_completos':
                        inscripciones = inscripciones.filter(preinscripcionbecarequisito__cumplerequisito=True)\
                            .annotate(total_requisitos=Count('preinscripcionbecarequisito', filter=Q(status=True), distinct=True))\
                            .filter(total_requisitos=cantidad_total_requisitos)
                    elif action_extra == 'requisitos_incompletos':
                        pendientes = inscripciones.values_list('id', flat=True).filter(preinscripcionbecarequisito__cumplerequisito__isnull=True)
                        inscripciones = inscripciones.filter(preinscripcionbecarequisito__cumplerequisito=False).exclude(id__in=pendientes).distinct()
                    matriculados = MatriculaGrupoSocioEconomico.objects.filter(
                        Q(gruposocioeconomico__codigo='D') | Q(gruposocioeconomico__codigo='C-'),
                        matricula__retiradomatricula=False,
                        matricula__nivel__periodo=periodosesion,
                        matricula__matriculagruposocioeconomico__tipomatricula=1,
                        matricula__inscripcion__in=inscripciones.values_list('inscripcion_id',flat=True),
                        matricula__inscripcion__carrera__coordinacion__excluir=False
                    ).distinct().order_by("gruposocioeconomico__codigo")
                row_num = 4
                i = 0
                for x in matriculados:
                    preinscripcion = PreInscripcionBeca.objects.filter(inscripcion=x.matricula.inscripcion, periodo=periodo, becatipo_id=18).first()
                    asignatura = MateriaAsignada.objects.filter(status=True,
                                                                matricula__inscripcion__id=x.matricula.inscripcion.id,
                                                                matricula__nivel__periodo__id=anterior.id,
                                                                materiaasignadaretiro__isnull=True)
                    verifica = 0
                    suma = 0
                    promedio = 0
                    sumasis = 0
                    total = asignatura.count()
                    for m in asignatura:
                        suma += m.notafinal
                        sumasis += m.asistenciafinal
                        if periodosesion.versionbeca is None or periodosesion.versionbeca == 1:
                            if m.estado.id != 1:
                                verifica = 1
                                break
                    if suma > 0 and sumasis > 0 and verifica == 0:
                        promedio = round(suma / total, 2)
                        asistencia = round(sumasis / total, 2)
                    if periodosesion.versionbeca > 1:
                        verifica = 0
                        promedio = x.matricula.inscripcion.promedio

                    #if verifica == 0:
                    campo1 = preinscripcion.id
                    campo2 = preinscripcion.orden
                    campo3 = x.matricula.inscripcion.persona.nombre_completo_inverso()
                    if (x.matricula.inscripcion.carrera.mencion and x.matricula.inscripcion.carrera.nombre != ''):
                        campo4 = x.matricula.inscripcion.carrera.nombre + ' CON MENCION EN  ' + x.matricula.inscripcion.carrera.mencion
                    elif (x.matricula.inscripcion.carrera.nombre != ''):
                        campo4 = x.matricula.inscripcion.carrera.nombre
                    campo5 = x.matricula.inscripcion.modalidad.__str__()
                    campo6 = x.matricula.inscripcion.persona.cedula
                    campo7 = x.matricula.inscripcion.persona.nacimiento
                    campo8 = x.matricula.inscripcion.persona.direccion_completa()
                    campo9 = promedio
                    campo10 = preinscripcion.promedio
                    campo11 = asistencia
                    campo12 = x.gruposocioeconomico.codigo
                    campo13 = x.gruposocioeconomico.nombre
                    campo14 = x.matricula.inscripcion.sesion.nombre
                    if x.matricula.nivelmalla.nombre:
                        campo15 = x.matricula.nivelmalla.nombre
                    else:
                        campo15 = 'Ninguno'
                    if x.matricula.paralelo:
                        campo16 = x.matricula.paralelo.nombre
                    else:
                        campo16 = 'Ninguno'
                    campo17 = x.matricula.inscripcion.persona.direccion_completa()
                    campo18 = x.matricula.inscripcion.persona.telefono
                    campo19 = x.matricula.inscripcion.persona.telefono_conv
                    campo20 = x.matricula.inscripcion.persona.email
                    campo21 = x.matricula.inscripcion.persona.emailinst
                    campo22 = x.matricula.inscripcion.persona.sexo.nombre if x.matricula.inscripcion.persona.sexo else ""
                    campo23 = x.matricula.inscripcion.persona.pais.nombre if x.matricula.inscripcion.persona.pais else ""
                    campo24 = x.matricula.inscripcion.persona.provincia.nombre if x.matricula.inscripcion.persona.provincia else ""
                    campo25 = x.matricula.inscripcion.persona.canton.nombre if x.matricula.inscripcion.persona.canton else ""
                    campo26 = x.matricula.inscripcion.persona.direccion if x.matricula.inscripcion.persona.direccion else ""
                    campo27 = x.matricula.inscripcion.persona.direccion2 if x.matricula.inscripcion.persona.direccion2 else ""
                    campo28 = x.matricula.inscripcion.persona.sector if x.matricula.inscripcion.persona.sector else ""
                    campo29 = x.matricula.estado_renovacion_beca(becatipo, anterior)
                    i += 1
                    ws.write(row_num, 0, i, font_style2)
                    ws.write(row_num, 1, campo1, font_style2)
                    ws.write(row_num, 2, campo2, font_style2)
                    ws.write(row_num, 3, campo3, font_style2)
                    ws.write(row_num, 4, campo4, font_style2)
                    ws.write(row_num, 5, campo5, date_format)
                    ws.write(row_num, 6, campo6, font_style2)
                    ws.write(row_num, 7, campo7, date_format)
                    ws.write(row_num, 8, campo8, font_style2)
                    ws.write(row_num, 9, campo9, font_style2)
                    ws.write(row_num, 10, campo10, font_style2)
                    ws.write(row_num, 11, campo11, font_style2)
                    ws.write(row_num, 12, campo12, font_style2)
                    ws.write(row_num, 13, campo13, font_style2)
                    ws.write(row_num, 14, campo14, font_style2)
                    ws.write(row_num, 15, campo15, font_style2)
                    ws.write(row_num, 16, campo16, font_style2)
                    ws.write(row_num, 17, campo17, font_style2)
                    ws.write(row_num, 18, campo18, font_style2)
                    ws.write(row_num, 19, campo19, font_style2)
                    ws.write(row_num, 20, campo20, font_style2)
                    ws.write(row_num, 21, campo21, font_style2)
                    ws.write(row_num, 22, campo22, font_style2)
                    ws.write(row_num, 23, campo23, font_style2)
                    ws.write(row_num, 24, campo24, font_style2)
                    ws.write(row_num, 25, campo25, font_style2)
                    ws.write(row_num, 26, campo26, font_style2)
                    ws.write(row_num, 27, campo27, font_style2)
                    ws.write(row_num, 28, campo28, font_style2)
                    ws.write(row_num, 29, str(periodoactual), font_style2)
                    ws.write(row_num, 30, str(periodovalida), font_style2)
                    ws.write(row_num, 31, campo29, font_style2)
                    if requisitos:
                        col_num = len(columns) - requisitos.count()*(2 if not action_extra else 1)
                        for detallerequisitobeca in requisitos:
                            preinscriocionrequisitos = PreInscripcionBecaRequisito.objects.filter(preinscripcion=preinscripcion, detallerequisitobeca=detallerequisitobeca, status=True)
                            for preinsrequisito in preinscriocionrequisitos:
                                cumple = ''
                                cumplerequisito = preinsrequisito.cumplerequisito
                                if cumplerequisito:
                                   cumple = 'SI'
                                elif cumplerequisito == False:
                                   cumple = 'NO'
                                ws.write(row_num, col_num, cumple, font_style2)
                                if not action_extra:
                                    col_num += 1
                                    ws.write(row_num, col_num, '', font_style2) #Observación por columna
                            col_num += 1



                    row_num += 1
                        # else:
                        #     verifica =123
                wb.save(response)
                return response
            except Exception as ex:
                err = 'Error on line {}'.format(sys.exc_info()[-1].tb_lineno)
                msg = ex.__str__()
                msg = f'{msg} {err}'

                return JsonResponse({"result": "bad", "mensaje": u"Error: %s"%msg})

        elif action == 'imprimir_promediook':
            try:
                becatipo = BecaTipo.objects.get(pk=17)
                configuracion = becatipo.becatipoconfiguracion_set.filter(becaperiodo__periodo=periodo).first()

                action_extra = request.POST.get('a')
                cantidad_total_requisitos = 0
                nombre_archivo = 'alumnos_promedio'
                titulo_archivo = 'LISTADO DE ESTUDIANTE PRESELECCIONADO POR ALTO PROMEDIO Y DISTINCIÓN ACADÉMICA'
                if configuracion:
                    if action_extra:
                        cantidad_total_requisitos = configuracion.requisitosbecas.count() if configuracion.requisitosbecas else 0
                        if action_extra == 'requisitos_completos':
                            nombre_archivo = 'alumnos_promedio_OCAS'
                            titulo_archivo = 'LISTADO DE ESTUDIANTES SELECCIONADOS POR ALTO PROMEDIO Y DISTINCIÓN ACADÉMICA'
                        elif action_extra == 'requisitos_incompletos':
                            nombre_archivo = 'alumnos_promedio_rechazados'
                            titulo_archivo = 'LISTADO DE ESTUDIANTES NO SELECCIONADOS POR ALTO PROMEDIO Y DISTINCIÓN ACADÉMICA'

                font_style = XFStyle()
                font_style.font.bold = True
                font_style2 = XFStyle()
                font_style2.font.bold = False
                wb = Workbook(encoding='utf-8')
                ws = wb.add_sheet('alumnos_promedio')
                response = HttpResponse(content_type="application/ms-excel")
                columns = [
                    (u"N.", 1500),
                    (u"ID_PREINSCRIPCION.", 1500),
                    (u"ORDEN.", 1000),
                    (u"APELLIDOS Y NOMBRES", 12000),
                    (u"CARRERA", 3000),
                    (u"MODALIDAD", 3000),
                    (u"CEDULA", 3000),
                    (u"FECHA NACIMIENTO", 3000),
                    (u"PROMEDIO", 2000),
                    (u"PROMEDIO CARRERA", 2000),
                    (u"DESVIACIÓN ESTANDAR", 3000),
                    (u"ASISTENCIA", 2000),
                    (u"AMERITA", 2000),
                    (u"SESION", 3000),
                    (u"NIVEL", 3000),
                    (u"PARALELO", 3000),
                    (u"DIRECCION COMPLETA", 4000),
                    (u"TELEFONO", 4000),
                    (u"TELEFONO CONVENCIONAL", 4000),
                    (u"EMAIL", 5000),
                    (u"EMAIL INSTITUCIONAL", 5000),
                    (u"SEXO", 5000),
                    (u"PAIS", 8000),
                    (u"PROVINCIA", 8000),
                    (u"CANTON", 8000),
                    (u"DIRECCION 1", 8000),
                    (u"DIRECCION 2", 8000),
                    (u"SECTOR", 8000),
                    (u"GRUPO SOCIO ECONÓMICO", 8000),
                    (u"TIPO/ESTADO", 2500),
                ]
                requisitos = []
                if configuracion:
                    #Requisitos de tipos de becas
                    requisitos = configuracion.requisitosbecas.filter(visible=True, status=True)
                    for detallerequisitobeca in requisitos:
                        columns.append((f'({detallerequisitobeca.pk}) {detallerequisitobeca.requisitobeca.__str__()}', 10000))
                        if not action_extra:
                            columns.append((f'({detallerequisitobeca.pk}) OBSERVACIÓN', 5000))

                response['Content-Disposition'] = f'attachment; filename={nombre_archivo}' + random.randint(1, 10000).__str__() + '.xls'
                style_title = xlwt.easyxf(
                    'font: height 350, name Arial, colour_index black, bold on, italic on; align: wrap on, vert centre, horiz center;')
                style_title_2 = xlwt.easyxf(
                    'font: height 250, name Arial, colour_index black, bold on, italic on; align: wrap on, vert centre, horiz center;')
                ws.write_merge(0, 0, 0, len(columns), 'UNIVERSIDAD ESTATAL DE MILAGRO', style_title)
                ws.write_merge(1, 1, 0, len(columns), titulo_archivo, style_title_2)
                row_num = 3
                for col_num in range(len(columns)):
                    ws.write(row_num, col_num, columns[col_num][0], font_style)
                    ws.col(col_num).width = columns[col_num][1]
                date_format = xlwt.XFStyle()
                date_format.num_format_str = 'yyyy/mm/dd'
                if periodosesion.versionbeca is None or periodosesion.versionbeca == 1:
                    inscripcionesactuales = Matricula.objects.values_list('inscripcion__id', flat=True).filter(
                        status=True, nivel__periodo__id=periodo.id,
                        estado_matricula__in=[2, 3],
                        retiradomatricula=False,
                        matriculagruposocioeconomico__tipomatricula=1).exclude(
                        inscripcion__persona__rubro__cancelado=False,
                        inscripcion__persona__rubro__status=True,
                        inscripcion__persona__rubro__fecha__lte=fechafinperiodonterior).distinct().order_by(
                        "inscripcion__persona")
                    matriculados = Matricula.objects.filter(status=True,
                                                            nivel__periodo__id=anterior.id,
                                                            estado_matricula__in=[2, 3],
                                                            retiradomatricula=False,
                                                            matriculagruposocioeconomico__tipomatricula=1,
                                                            inscripcion__id__in=inscripcionesactuales).exclude(
                        inscripcion__persona__rubro__cancelado=False,
                        inscripcion__persona__rubro__status=True,
                        inscripcion__persona__rubro__fecha__lte=fechafinperiodonterior).distinct().order_by(
                        "inscripcion__persona")
                elif periodosesion.versionbeca == 2:
                    inscripciones = PreInscripcionBeca.objects.filter(periodo=periodo, becatipo_id=17).values_list('inscripcion_id', flat=True)

                    if action_extra == 'requisitos_completos':
                        inscripciones = inscripciones.filter(preinscripcionbecarequisito__cumplerequisito=True)\
                            .annotate(total_requisitos=Count('preinscripcionbecarequisito', filter=Q(status=True), distinct=True))\
                            .filter(total_requisitos=cantidad_total_requisitos)
                    elif action_extra == 'requisitos_incompletos':
                        pendientes = inscripciones.values_list('id', flat=True).filter(preinscripcionbecarequisito__cumplerequisito__isnull=True)
                        inscripciones = inscripciones.filter(preinscripcionbecarequisito__cumplerequisito=False).exclude(id__in=pendientes).distinct()

                    matriculados = Matricula.objects.filter(inscripcion__id__in=inscripciones.values_list('inscripcion_id',flat=True),
                                                            nivel__periodo__id=periodo.id).distinct().order_by("inscripcion__persona")

                row_num = 4
                i = 0
                continua = False
                for x in matriculados:
                    preinscripcion = PreInscripcionBeca.objects.filter(inscripcion=x.inscripcion, periodo=periodo, becatipo_id=17).first()
                    amerita = None
                    verifica = 0
                    suma = 0
                    sumasis = 0
                    promedio = 0

                    materias = MateriaAsignada.objects.filter(status=True,
                                                              matricula__inscripcion__id=x.inscripcion.id,
                                                              matricula__nivel__periodo__id=anterior.id,
                                                              materiaasignadaretiro__isnull=True)
                    total = materias.count()
                    if total < 4:
                        amerita = "<4: " + str(total)
                    for m in materias:
                        suma += m.notafinal
                        sumasis += m.asistenciafinal
                        if m.estado.id != 1:
                            verifica = 1
                            break
                    if suma > 0 and sumasis > 0:
                        promedio = round(suma / total, 2)
                        asistencia = round(sumasis / total, 2)

                    if periodosesion.versionbeca is None or periodosesion.versionbeca == 1:
                        if verifica == 0 and promedio >= 85:
                            continua = True
                    elif periodosesion.versionbeca == 2:
                        continua = True
                        promedio = x.inscripcion.promedio
                    #if continua:
                    campo1 = preinscripcion.id
                    campo2 = preinscripcion.orden
                    campo3 = x.inscripcion.persona.nombre_completo_inverso()
                    if x.inscripcion.carrera.mencion:
                        campo4 = x.inscripcion.carrera.nombre + ' CON MENCION EN  ' + x.inscripcion.carrera.mencion
                    else:
                        campo4 = x.inscripcion.carrera.nombre
                    campo5 = x.inscripcion.modalidad.__str__()
                    campo6 = x.inscripcion.persona.cedula
                    campo7 = x.inscripcion.persona.nacimiento
                    campo8 = preinscripcion.promedio
                    campo9 = preinscripcion.promedio_carrera
                    campo10 = preinscripcion.desviacion_estandar
                    campo11 = asistencia
                    campo12 = amerita
                    campo13 = x.inscripcion.sesion.nombre
                    campo14 = x.nivelmalla.nombre
                    if x.paralelo:
                        campo15 = x.paralelo.nombre
                    else:
                        campo15 = 'Ninguno'
                    campo16 = x.inscripcion.persona.direccion_completa()
                    campo17 = x.inscripcion.persona.telefono
                    campo18 = x.inscripcion.persona.telefono_conv
                    campo19 = x.inscripcion.persona.email
                    campo20 = x.inscripcion.persona.emailinst
                    campo21 = x.inscripcion.persona.sexo.nombre
                    campo22 = x.inscripcion.persona.pais.nombre if x.inscripcion.persona.pais else ""
                    campo23 = x.inscripcion.persona.provincia.nombre if x.inscripcion.persona.provincia else ""
                    campo24 = x.inscripcion.persona.canton.nombre if x.inscripcion.persona.canton else ""
                    campo25 = x.inscripcion.persona.direccion if x.inscripcion.persona.direccion else ""
                    campo26 = x.inscripcion.persona.direccion2 if x.inscripcion.persona.direccion2 else ""
                    campo27 = x.inscripcion.persona.sector if x.inscripcion.persona.sector else ""
                    campo28 = str(
                        x.matriculagruposocioeconomico().nombre) if x.matriculagruposocioeconomico() else ""
                    campo29 = x.estado_renovacion_beca(becatipo, anterior)
                    i += 1
                    ws.write(row_num, 0, i, font_style2)
                    ws.write(row_num, 1, campo1, font_style2)
                    ws.write(row_num, 2, campo2, font_style2)
                    ws.write(row_num, 3, campo3, font_style2)
                    ws.write(row_num, 4, campo4, font_style2)
                    ws.write(row_num, 5, campo5, font_style2)
                    ws.write(row_num, 6, campo6, font_style2)
                    ws.write(row_num, 7, campo7, date_format)
                    ws.write(row_num, 8, campo8, font_style2)
                    ws.write(row_num, 9, campo9, font_style2)
                    ws.write(row_num, 10, campo10, font_style2)
                    ws.write(row_num, 11, campo11, font_style2)
                    ws.write(row_num, 12, campo12, font_style2)
                    ws.write(row_num, 13, campo13, font_style2)
                    ws.write(row_num, 14, campo14, font_style2)
                    ws.write(row_num, 15, campo15, font_style2)
                    ws.write(row_num, 16, campo16, font_style2)
                    ws.write(row_num, 17, campo17, font_style2)
                    ws.write(row_num, 18, campo18, font_style2)
                    ws.write(row_num, 19, campo19, font_style2)
                    ws.write(row_num, 20, campo20, font_style2)
                    ws.write(row_num, 21, campo21, font_style2)
                    ws.write(row_num, 22, campo22, font_style2)
                    ws.write(row_num, 23, campo23, font_style2)
                    ws.write(row_num, 24, campo24, font_style2)
                    ws.write(row_num, 25, campo25, font_style2)
                    ws.write(row_num, 26, campo26, font_style2)
                    ws.write(row_num, 27, campo27, font_style2)
                    ws.write(row_num, 28, campo28, font_style2)
                    ws.write(row_num, 29, campo29, font_style2)
                    if requisitos:
                        col_num = len(columns) - requisitos.count()*(2 if not action_extra else 1)
                        for detallerequisitobeca in requisitos:
                            preinscriocionrequisitos = PreInscripcionBecaRequisito.objects.filter(preinscripcion=preinscripcion, detallerequisitobeca=detallerequisitobeca, status=True)
                            for preinsrequisito in preinscriocionrequisitos:
                                cumple = ''
                                cumplerequisito = preinsrequisito.cumplerequisito
                                if cumplerequisito:
                                   cumple = 'SI'
                                elif cumplerequisito == False:
                                   cumple = 'NO'
                                ws.write(row_num, col_num, cumple, font_style2)
                                if not action_extra:
                                    col_num += 1
                                    ws.write(row_num, col_num, '', font_style2) #Observación por columna
                            col_num += 1
                    row_num += 1
                wb.save(response)
                return response
            except Exception as ex:
                err = 'Error on line {}'.format(sys.exc_info()[-1].tb_lineno)
                msg = ex.__str__()
                msg = f'{msg} {err}'

                return JsonResponse({"result": "bad", "mensaje": u"Error: %s"%msg})

        elif action == 'imprimir_primernivel':
            try:
                becatipo = BecaTipo.objects.get(pk=16)
                configuracion = becatipo.becatipoconfiguracion_set.filter(becaperiodo__periodo=periodo).first()

                action_extra = request.POST.get('a')
                cantidad_total_requisitos = 0
                nombre_archivo = 'alumnos_primernivel'
                titulo_archivo = 'LISTADO DE ESTUDIANTE PRESELECCIONADO POR PRIMER NIVEL'
                if configuracion:
                    if action_extra:
                        cantidad_total_requisitos = configuracion.requisitosbecas.count() if configuracion.requisitosbecas else 0
                        if action_extra == 'requisitos_completos':
                            nombre_archivo = 'alumnos_primernivel_OCAS'
                            titulo_archivo = 'LISTADO DE ESTUDIANTES SELECCIONADOS POR PRIMER NIVEL'
                        elif action_extra == 'requisitos_incompletos':
                            nombre_archivo = 'alumnos_primernivel_rechazados'
                            titulo_archivo = 'LISTADO DE ESTUDIANTES NO SELECCIONADOS POR PRIMER NIVEL'
                font_style = XFStyle()
                font_style.font.bold = True
                font_style2 = XFStyle()
                font_style2.font.bold = False
                wb = Workbook(encoding='utf-8')
                ws = wb.add_sheet('alumnos_primernivel')
                response = HttpResponse(content_type="application/ms-excel")
                columns = [
                    (u"N.", 1500),
                    (u"ID_PREINSCRIPCION.", 1500),
                    (u"ORDEN", 1000),
                    (u"APELLIDOS Y NOMBRES", 12000),
                    (u"CARRERA", 3000),
                    (u"MODALIDAD", 3000),
                    (u"CEDULA", 3000),
                    (u"FECHA NACIMIENTO", 3000),
                    (u"SESION", 3000),
                    (u"NIVEL", 3000),
                    (u"PARALELO", 3000),
                    (u"DIRECCION COMPLETA", 4000),
                    (u"TELEFONO", 4000),
                    (u"TELEFONO CONVENCIONAL", 4000),
                    (u"EMAIL", 5000),
                    (u"EMAIL INSTITUCIONAL", 5000),
                    (u"SEXO", 5000),
                    (u"PAIS", 8000),
                    (u"PROVINCIA", 8000),
                    (u"CANTON", 8000),
                    (u"DIRECCION 1", 8000),
                    (u"DIRECCION 2", 8000),
                    (u"SECTOR", 8000),
                    (u"GRUPO SOCIO ECONÓMICO", 8000),
                    (u"TÍTULO", 20000),
                    (u"COLEGIO", 20000),
                    (u"FECHA INICIO", 3000),
                    (u"FECHA OBTENCIÓN", 3000),
                    (u"FECHA EGRESADO", 3000),
                    (u"CALIFICACIÓN", 2000),
                    (u"AÑO INICIO", 2000),
                    (u"AÑO FIN", 2000),
                    (u"ACTA DE GRADO", 20000),
                ]


                requisitos = []
                if configuracion:
                    #Requisitos de tipos de becas
                    requisitos = configuracion.requisitosbecas.filter(visible=True, status=True)
                    for detallerequisitobeca in requisitos:
                        columns.append((f'({detallerequisitobeca.pk}) {detallerequisitobeca.requisitobeca.__str__()}', 10000))
                        if not action_extra:
                            columns.append((f'({detallerequisitobeca.pk}) OBSERVACIÓN', 5000))

                response['Content-Disposition'] = f'attachment; filename={nombre_archivo}' + random.randint(1, 10000).__str__() + '.xls'
                style_title = xlwt.easyxf(
                    'font: height 350, name Arial, colour_index black, bold on, italic on; align: wrap on, vert centre, horiz center;')
                style_title_2 = xlwt.easyxf(
                    'font: height 250, name Arial, colour_index black, bold on, italic on; align: wrap on, vert centre, horiz center;')
                ws.write_merge(0, 0, 0, len(columns), 'UNIVERSIDAD ESTATAL DE MILAGRO', style_title)
                ws.write_merge(1, 1, 0, len(columns), titulo_archivo, style_title_2)
                row_num = 3
                for col_num in range(len(columns)):
                    ws.write(row_num, col_num, columns[col_num][0], font_style)
                    ws.col(col_num).width = columns[col_num][1]
                date_format = xlwt.XFStyle()
                date_format.num_format_str = 'yyyy/mm/dd'
                preinscripciones = PreInscripcionBeca.objects.filter(periodo=periodo,
                                                                     becatipo=becatipo,
                                                                     inscripcion__matricula__estado_matricula__in=[2, 3],
                                                                     status=True,
                                                                     inscripcion__matricula__status=True,
                                                                     inscripcion__matricula__retiradomatricula=False,
                                                                     inscripcion__matricula__matriculagruposocioeconomico__tipomatricula=1).distinct()

                if action_extra == 'requisitos_completos':
                    preinscripciones = preinscripciones.filter(preinscripcionbecarequisito__cumplerequisito=True) \
                        .annotate(total_requisitos=Count('preinscripcionbecarequisito', filter=Q(status=True), distinct=True)) \
                        .filter(total_requisitos=cantidad_total_requisitos)
                elif action_extra == 'requisitos_incompletos':
                    pendientes = preinscripciones.values_list('id', flat=True).filter(preinscripcionbecarequisito__cumplerequisito__isnull=True)
                    preinscripciones = preinscripciones.filter(preinscripcionbecarequisito__cumplerequisito=False).exclude(id__in=pendientes).distinct()

                row_num = 4

                for i, preinscripcion in enumerate(preinscripciones):
                    campo1 = preinscripcion.id
                    campo2 = preinscripcion.orden
                    campo3 = preinscripcion.inscripcion.persona.nombre_completo_inverso()
                    if preinscripcion.inscripcion.carrera.mencion:
                        campo4 = preinscripcion.inscripcion.carrera.nombre + ' CON MENCION EN  ' + preinscripcion.inscripcion.carrera.mencion
                    else:
                        campo4 = preinscripcion.inscripcion.carrera.nombre
                    campo5 = preinscripcion.inscripcion.modalidad.__str__()
                    campo6 = preinscripcion.inscripcion.persona.cedula
                    campo7 = preinscripcion.inscripcion.persona.nacimiento
                    campo8 = preinscripcion.inscripcion.sesion.nombre
                    matricula = preinscripcion.inscripcion.matricula_set.filter(nivel__periodo=periodo, retiradomatricula=False, status=True).first()
                    campo9 = matricula.nivelmalla.nombre
                    campo10 = matricula.paralelo.nombre if matricula.paralelo else "Ninguno"
                    campo11 = preinscripcion.inscripcion.persona.direccion_completa()
                    campo12 = preinscripcion.inscripcion.persona.telefono
                    campo13 = preinscripcion.inscripcion.persona.telefono_conv
                    campo14 = preinscripcion.inscripcion.persona.email
                    campo15 = preinscripcion.inscripcion.persona.emailinst
                    campo16 = preinscripcion.inscripcion.persona.sexo.nombre
                    campo17 = preinscripcion.inscripcion.persona.pais.nombre if preinscripcion.inscripcion.persona.pais else ""
                    campo18 = preinscripcion.inscripcion.persona.provincia.nombre if preinscripcion.inscripcion.persona.provincia else ""
                    campo19 = preinscripcion.inscripcion.persona.canton.nombre if preinscripcion.inscripcion.persona.canton else ""
                    campo20 = preinscripcion.inscripcion.persona.direccion if preinscripcion.inscripcion.persona.direccion else ""
                    campo21 = preinscripcion.inscripcion.persona.direccion2 if preinscripcion.inscripcion.persona.direccion2 else ""
                    campo22 = preinscripcion.inscripcion.persona.sector if preinscripcion.inscripcion.persona.sector else ""
                    campo23 = str(matricula.matriculagruposocioeconomico().nombre) if matricula.matriculagruposocioeconomico() else ""
                    campo24 = ""
                    campo25 = ""
                    campo26 = ""
                    campo27 = ""
                    campo28 = ""
                    campo29 = ""
                    campo30 = ""
                    campo31 = ""
                    campo32 = ""
                    titulo_bachiller = preinscripcion.inscripcion.persona.titulo_bachiller()
                    if titulo_bachiller:
                        campo24 = titulo_bachiller.titulo.__str__() #titulo
                        campo25 = titulo_bachiller.colegio.nombre if titulo_bachiller.colegio else "Sin Colegio" #colegio
                        campo26 = titulo_bachiller.fechainicio #fechainicio
                        campo27 = titulo_bachiller.fechaobtencion #fechaobtencion
                        campo28 = titulo_bachiller.fechaegresado #fechaegresado
                        detalle = titulo_bachiller.detalletitulacion()
                        if detalle:
                            campo29 = detalle.calificacion #calificacion
                            campo30 = detalle.anioinicioperiodograduacion #anioinicio
                            campo31 = detalle.aniofinperiodograduacion #aniofin
                            campo32 = detalle.actagrado.url if detalle.actagrado else "" #actagrado
                    ws.write(row_num, 0, i+1, font_style2)
                    ws.write(row_num, 1, campo1, font_style2)
                    ws.write(row_num, 2, campo2, font_style2)
                    ws.write(row_num, 3, campo3, font_style2)
                    ws.write(row_num, 4, campo4, font_style2)
                    ws.write(row_num, 5, campo5, font_style2)
                    ws.write(row_num, 6, campo6, font_style2)
                    ws.write(row_num, 7, campo7, date_format)
                    ws.write(row_num, 8, campo8, font_style2)
                    ws.write(row_num, 9, campo9, font_style2)
                    ws.write(row_num, 10, campo10, font_style2)
                    ws.write(row_num, 11, campo11, font_style2)
                    ws.write(row_num, 12, campo12, font_style2)
                    ws.write(row_num, 13, campo13, font_style2)
                    ws.write(row_num, 14, campo14, font_style2)
                    ws.write(row_num, 15, campo15, font_style2)
                    ws.write(row_num, 16, campo16, font_style2)
                    ws.write(row_num, 17, campo17, font_style2)
                    ws.write(row_num, 18, campo18, font_style2)
                    ws.write(row_num, 19, campo19, font_style2)
                    ws.write(row_num, 20, campo20, font_style2)
                    ws.write(row_num, 21, campo21, font_style2)
                    ws.write(row_num, 22, campo22, font_style2)
                    ws.write(row_num, 23, campo23, font_style2)
                    ws.write(row_num, 24, campo24, font_style2)
                    ws.write(row_num, 25, campo25, font_style2)
                    ws.write(row_num, 26, campo26, date_format)
                    ws.write(row_num, 27, campo27, date_format)
                    ws.write(row_num, 28, campo28, date_format)
                    ws.write(row_num, 29, campo29, font_style2)
                    ws.write(row_num, 30, campo30, font_style2)
                    ws.write(row_num, 31, campo31, font_style2)
                    ws.write(row_num, 32, campo32, font_style2)

                    if requisitos:
                        col_num = len(columns) - requisitos.count()*(2 if not action_extra else 1)
                        for detallerequisitobeca in requisitos:
                            preinscriocionrequisitos = PreInscripcionBecaRequisito.objects.filter(preinscripcion=preinscripcion, detallerequisitobeca=detallerequisitobeca, status=True)
                            for preinsrequisito in preinscriocionrequisitos:
                                cumple = ''
                                cumplerequisito = preinsrequisito.cumplerequisito
                                if cumplerequisito:
                                    cumple = 'SI'
                                elif cumplerequisito == False:
                                    cumple = 'NO'
                                ws.write(row_num, col_num, cumple, font_style2)
                                if not action_extra:
                                    col_num += 1
                                    ws.write(row_num, col_num, '', font_style2)  # Observación por columna
                            col_num += 1
                    row_num += 1
                wb.save(response)
                return response
            except Exception as ex:
                err = 'Error on line {}'.format(sys.exc_info()[-1].tb_lineno)
                msg = ex.__str__()
                msg = f'{msg} {err}'

                return JsonResponse({"result": "bad", "mensaje": u"Error: %s"%msg})

        elif action == 'imprimir_deporte':
            try:
                becatipo = BecaTipo.objects.get(pk=20)
                configuracion = becatipo.becatipoconfiguracion_set.filter(becaperiodo__periodo=periodo).first()

                action_extra = request.POST.get('a')
                cantidad_total_requisitos = 0
                nombre_archivo = 'alumnos_deporte'
                titulo_archivo = 'LISTADO DE ESTUDIANTE PRESELECCIONADO POR ALTO RENDIMIENTO EN DEPORTES'
                if configuracion:
                    if action_extra:
                        cantidad_total_requisitos = configuracion.requisitosbecas.count() if configuracion.requisitosbecas else 0
                        if action_extra == 'requisitos_completos':
                            nombre_archivo = 'alumnos_deporte_OCAS'
                            titulo_archivo = 'LISTADO DE ESTUDIANTES SELECCIONADOS POR ALTO RENDIMIENTO EN DEPORTES'
                        elif action_extra == 'requisitos_incompletos':
                            nombre_archivo = 'alumnos_deporte_rechazados'
                            titulo_archivo = 'LISTADO DE ESTUDIANTES NO SELECCIONADOS POR ALTO RENDIMIENTO EN DEPORTES'
                font_style = XFStyle()
                font_style.font.bold = True
                font_style2 = XFStyle()
                font_style2.font.bold = False
                wb = Workbook(encoding='utf-8')
                ws = wb.add_sheet('alumnos_deporte')
                response = HttpResponse(content_type="application/ms-excel")
                columns = [
                    (u"N.", 1500),
                    (u"ID_PREINSCRIPCION.", 1500),
                    (u"ORDEN", 1500),
                    (u"APELLIDOS Y NOMBRES", 12000),
                    (u"CARRERA", 3000),
                    (u"MODALIDAD", 3000),
                    (u"CEDULA", 3000),
                    (u"FECHA NACIMIENTO", 3000),
                    (u"PROMEDIO", 2000),
                    (u"ASISTENCIA", 2000),
                    (u"AMERITA", 2000),
                    (u"SESION", 3000),
                    (u"NIVEL", 3000),
                    (u"PARALELO", 3000),
                    (u"DIRECCION COMPLETA", 4000),
                    (u"TELEFONO", 4000),
                    (u"TELEFONO CONVENCIONAL", 4000),
                    (u"EMAIL", 5000),
                    (u"EMAIL INSTITUCIONAL", 5000),
                    (u"SEXO", 5000),
                    (u"PAIS", 8000),
                    (u"PROVINCIA", 8000),
                    (u"CANTON", 8000),
                    (u"DIRECCION 1", 8000),
                    (u"DIRECCION 2", 8000),
                    (u"SECTOR", 8000),
                    (u"GRUPO SOCIO ECONÓMICO", 8000),
                    (u"TIPO/ESTADO", 2500),
                ]

                requisitos = []
                if configuracion:
                    #Requisitos de tipos de becas
                    requisitos = configuracion.requisitosbecas.filter(visible=True, status=True)
                    for detallerequisitobeca in requisitos:
                        columns.append((f'({detallerequisitobeca.pk}) {detallerequisitobeca.requisitobeca.__str__()}', 10000))
                        if not action_extra:
                            columns.append((f'({detallerequisitobeca.pk}) OBSERVACIÓN', 5000))

                response['Content-Disposition'] = f'attachment; filename={nombre_archivo}' + random.randint(1, 10000).__str__() + '.xls'
                style_title = xlwt.easyxf(
                    'font: height 350, name Arial, colour_index black, bold on, italic on; align: wrap on, vert centre, horiz center;')
                style_title_2 = xlwt.easyxf(
                    'font: height 250, name Arial, colour_index black, bold on, italic on; align: wrap on, vert centre, horiz center;')
                ws.write_merge(0, 0, 0, len(columns), 'UNIVERSIDAD ESTATAL DE MILAGRO', style_title)
                ws.write_merge(1, 1, 0, len(columns),titulo_archivo, style_title_2)
                row_num = 3
                for col_num in range(len(columns)):
                    ws.write(row_num, col_num, columns[col_num][0], font_style)
                    ws.col(col_num).width = columns[col_num][1]
                date_format = xlwt.XFStyle()
                date_format.num_format_str = 'yyyy/mm/dd'
                if periodosesion.versionbeca is None or periodosesion.versionbeca == 1:
                    matriculados = Matricula.objects.filter(pk=None)
                elif periodosesion.versionbeca == 2:
                    inscripciones = PreInscripcionBeca.objects.filter(periodo=periodo, becatipo_id=20).values_list(
                        'inscripcion_id', flat=True)

                    if action_extra == 'requisitos_completos':
                        inscripciones = inscripciones.filter(preinscripcionbecarequisito__cumplerequisito=True) \
                            .annotate(total_requisitos=Count('preinscripcionbecarequisito', filter=Q(status=True), distinct=True)) \
                            .filter(total_requisitos=cantidad_total_requisitos)
                    elif action_extra == 'requisitos_incompletos':
                        pendientes = inscripciones.values_list('id', flat=True).filter(
                            preinscripcionbecarequisito__cumplerequisito__isnull=True)
                        inscripciones = inscripciones.filter(
                            preinscripcionbecarequisito__cumplerequisito=False).exclude(
                            id__in=pendientes).distinct()

                    matriculados = Matricula.objects.filter(inscripcion__id__in=inscripciones.values_list('inscripcion_id', flat=True),
                                                            nivel__periodo__id=periodo.id).distinct().order_by("inscripcion__persona")
                row_num = 4
                i = 0
                continua = False
                for x in matriculados:
                    preinscripcion = PreInscripcionBeca.objects.filter(inscripcion=x.inscripcion, periodo=periodo, becatipo_id=20).first()
                    amerita = None
                    verifica = 0
                    suma = 0
                    sumasis = 0
                    promedio = 0
                    materias = MateriaAsignada.objects.filter(status=True,
                                                              matricula__inscripcion__id=x.inscripcion.id,
                                                              matricula__nivel__periodo__id=anterior.id,
                                                              materiaasignadaretiro__isnull=True)
                    total = materias.count()
                    if total < 4:
                        amerita = "<4: " + str(total)
                    for m in materias:
                        suma += m.notafinal
                        sumasis += m.asistenciafinal
                        if periodosesion.versionbeca is None or periodosesion.versionbeca == 1:
                            if m.estado.id != 1:
                                verifica = 1
                                break
                    if suma > 0 and sumasis > 0:
                        promedio = round(suma / total, 2)
                        asistencia = round(sumasis / total, 2)

                    if periodosesion.versionbeca is None or periodosesion.versionbeca == 1:
                        if verifica == 0 and promedio >= 85:
                            continua = True
                    elif periodosesion.versionbeca == 2:
                        if verifica == 0 and promedio >= 70:
                            continua = True
                    #if continua:
                    campo1 = preinscripcion.id
                    campo2 = preinscripcion.orden
                    campo3 = x.inscripcion.persona.nombre_completo_inverso()
                    if x.inscripcion.carrera.mencion:
                        campo4 = x.inscripcion.carrera.nombre + ' CON MENCION EN  ' + x.inscripcion.carrera.mencion
                    else:
                        campo4 = x.inscripcion.carrera.nombre
                    campo5 = x.inscripcion.modalidad.__str__()
                    campo6 = x.inscripcion.persona.cedula
                    campo7 = x.inscripcion.persona.nacimiento
                    campo8 = promedio
                    campo9 = asistencia
                    campo10 = amerita
                    campo11 = x.inscripcion.sesion.nombre
                    campo12 = x.nivelmalla.nombre
                    if x.paralelo:
                        campo13 = x.paralelo.nombre
                    else:
                        campo13 = 'Ninguno'
                    campo14 = x.inscripcion.persona.direccion_completa()
                    campo15 = x.inscripcion.persona.telefono
                    campo16 = x.inscripcion.persona.telefono_conv
                    campo17 = x.inscripcion.persona.email
                    campo18 = x.inscripcion.persona.emailinst
                    campo19 = x.inscripcion.persona.sexo.nombre
                    campo20 = x.inscripcion.persona.pais.nombre if x.inscripcion.persona.pais else ""
                    campo21 = x.inscripcion.persona.provincia.nombre if x.inscripcion.persona.provincia else ""
                    campo22 = x.inscripcion.persona.canton.nombre if x.inscripcion.persona.canton else ""
                    campo23 = x.inscripcion.persona.direccion if x.inscripcion.persona.direccion else ""
                    campo24 = x.inscripcion.persona.direccion2 if x.inscripcion.persona.direccion2 else ""
                    campo25 = x.inscripcion.persona.sector if x.inscripcion.persona.sector else ""
                    campo26 = str(
                        x.matriculagruposocioeconomico().nombre) if x.matriculagruposocioeconomico() else ""
                    campo27 = x.estado_renovacion_beca(becatipo, anterior)
                    i += 1
                    ws.write(row_num, 0, i, font_style2)
                    ws.write(row_num, 1, campo1, font_style2)
                    ws.write(row_num, 2, campo2, font_style2)
                    ws.write(row_num, 3, campo3, font_style2)
                    ws.write(row_num, 4, campo4, font_style2)
                    ws.write(row_num, 5, campo5, font_style2)
                    ws.write(row_num, 6, campo6, font_style2)
                    ws.write(row_num, 7, campo7, date_format)
                    ws.write(row_num, 8, campo8, font_style2)
                    ws.write(row_num, 9, campo9, font_style2)
                    ws.write(row_num, 10, campo10, font_style2)
                    ws.write(row_num, 11, campo11, font_style2)
                    ws.write(row_num, 12, campo12, font_style2)
                    ws.write(row_num, 13, campo13, font_style2)
                    ws.write(row_num, 14, campo14, font_style2)
                    ws.write(row_num, 15, campo15, font_style2)
                    ws.write(row_num, 16, campo16, font_style2)
                    ws.write(row_num, 17, campo17, font_style2)
                    ws.write(row_num, 18, campo18, font_style2)
                    ws.write(row_num, 19, campo19, font_style2)
                    ws.write(row_num, 20, campo20, font_style2)
                    ws.write(row_num, 21, campo21, font_style2)
                    ws.write(row_num, 22, campo22, font_style2)
                    ws.write(row_num, 23, campo23, font_style2)
                    ws.write(row_num, 24, campo24, font_style2)
                    ws.write(row_num, 25, campo25, font_style2)
                    ws.write(row_num, 26, campo26, font_style2)
                    ws.write(row_num, 27, campo27, font_style2)

                    if requisitos:
                        col_num = len(columns) - requisitos.count()*(2 if not action_extra else 1)
                        for detallerequisitobeca in requisitos:
                            preinscriocionrequisitos = PreInscripcionBecaRequisito.objects.filter(preinscripcion=preinscripcion, detallerequisitobeca=detallerequisitobeca, status=True)
                            for preinsrequisito in preinscriocionrequisitos:
                                cumple = ''
                                cumplerequisito = preinsrequisito.cumplerequisito
                                if cumplerequisito:
                                   cumple = 'SI'
                                elif cumplerequisito == False:
                                   cumple = 'NO'
                                ws.write(row_num, col_num, cumple, font_style2)
                                if not action_extra:
                                    col_num += 1
                                    ws.write(row_num, col_num, '', font_style2) #Observación por columna
                            col_num += 1
                    row_num += 1
                wb.save(response)
                return response
            except Exception as ex:
                err = 'Error on line {}'.format(sys.exc_info()[-1].tb_lineno)
                msg = ex.__str__()
                msg = f'{msg} {err}'

                return JsonResponse({"result": "bad", "mensaje": u"Error: %s"%msg})

        elif action == 'imprimir_promediodeuda':
            try:
                font_style = XFStyle()
                font_style.font.bold = True
                font_style2 = XFStyle()
                font_style2.font.bold = False
                wb = Workbook(encoding='utf-8')
                periodoactual = Periodo.objects.get(id=int(request.POST['periodoactual']))
                periodovalida = Periodo.objects.get(id=int(request.POST['periodovalida']))
                modalidad = int(request.POST['modalidad'])
                ws = wb.add_sheet('alumnos_promedio_condeuda')
                response = HttpResponse(content_type="application/ms-excel")
                response['Content-Disposition'] = 'attachment; filename=alumnos_promedio_condeuda' + random.randint(1,
                                                                                                                    10000).__str__() + '.xls'
                columns = [
                    (u"N.", 1500),
                    (u"APELLIDOS Y NOMBRES", 12000),
                    (u"CARRERA", 3000),
                    (u"CEDULA", 3000),
                    (u"PROMEDIO", 2000),
                    (u"ASISTENCIA", 2000),
                    (u"AMERITA", 2000),
                    (u"SESION", 3000),
                    (u"NIVEL", 3000),
                    (u"PARALELO", 3000),
                    (u"TELEFONO", 4000),
                    (u"TELEFONO CONVENCIONAL", 4000),
                    (u"EMAIL", 5000),
                    (u"EMAIL INSTITUCIONAL", 5000),
                    (u"SEXO", 5000),
                    (u"PAIS", 8000),
                    (u"PROVINCIA", 8000),
                    (u"CANTON", 8000),
                    (u"DIRECCION 1", 8000),
                    (u"DIRECCION 2", 8000),
                    (u"SECTOR", 8000),
                    (u"GRUPO SOCIO ECONÓMICO", 8000),
                    (u"PERIODO ACTUAL", 8000),
                    (u"PERIDOO VALIDA", 8000),
                ]
                row_num = 0
                for col_num in range(len(columns)):
                    ws.write(row_num, col_num, columns[col_num][0], font_style)
                    ws.col(col_num).width = columns[col_num][1]
                date_format = xlwt.XFStyle()
                date_format.num_format_str = 'yyyy/mm/dd'
                inscripcionesactuales = Matricula.objects.values_list('inscripcion__id', flat=True).filter(status=True,
                                                                                                           nivel__periodo=periodoactual,
                                                                                                           estado_matricula__in=[
                                                                                                               2, 3],
                                                                                                           retiradomatricula=False,
                                                                                                           matriculagruposocioeconomico__tipomatricula=1,
                                                                                                           inscripcion__carrera__modalidad=modalidad).exclude(
                    inscripcion__carrera__coordinacion__id__in=[7, 9]).distinct().order_by("inscripcion__persona")
                nuevolistado = Matricula.objects.filter(status=True,
                                                        nivel__periodo=periodovalida,
                                                        estado_matricula__in=[2, 3],
                                                        retiradomatricula=False,
                                                        matriculagruposocioeconomico__tipomatricula=1,
                                                        inscripcion__id__in=inscripcionesactuales,
                                                        inscripcion__carrera__modalidad=modalidad).exclude(
                    inscripcion__carrera__coordinacion__id__in=[7, 9]).distinct().order_by("inscripcion__persona")
                row_num = 1
                i = 0
                for x in nuevolistado:
                    amerita = None
                    materias = MateriaAsignada.objects.filter(matricula__status=True,
                                                              matricula__inscripcion=x.inscripcion,
                                                              matricula__nivel__periodo=periodovalida,
                                                              materiaasignadaretiro__isnull=True)
                    verifica = 0
                    suma = 0
                    sumasis = 0
                    promedio = 0
                    total = materias.count()
                    if total < 4:
                        amerita = "<4: " + str(total)
                    for m in materias:
                        suma += m.notafinal
                        sumasis += m.asistenciafinal
                        if m.estado.id != 1:
                            verifica = 1
                            break
                    if suma > 0 and sumasis > 0:
                        promedio = round(suma / total, 2)
                        asistencia = round(sumasis / total, 2)
                    if verifica == 0 and promedio >= 85:
                        campo1 = x.inscripcion.persona.nombre_completo_inverso()
                        if (x.inscripcion.carrera.mencion):
                            campo2 = x.inscripcion.carrera.nombre + ' CON MENCION EN  ' + x.inscripcion.carrera.mencion
                        else:
                            campo2 = x.inscripcion.carrera.nombre
                        campo3 = x.inscripcion.persona.cedula
                        campo4 = promedio
                        campo5 = asistencia
                        campo6 = amerita
                        campo7 = x.inscripcion.sesion.nombre
                        campo8 = x.nivelmalla.nombre
                        if (x.paralelo):
                            campo9 = x.paralelo.nombre
                        else:
                            campo9 = 'Ninguno'
                        campo11 = x.inscripcion.persona.telefono
                        campo12 = x.inscripcion.persona.telefono_conv
                        campo13 = x.inscripcion.persona.email
                        campo14 = x.inscripcion.persona.emailinst
                        campo15 = x.inscripcion.persona.sexo.nombre if x.inscripcion.persona.sexo else ""
                        campo10 = x.inscripcion.persona.pais.nombre if x.inscripcion.persona.pais else ""
                        campo16 = x.inscripcion.persona.provincia.nombre if x.inscripcion.persona.provincia else ""
                        campo17 = x.inscripcion.persona.canton.nombre if x.inscripcion.persona.canton else ""
                        campo18 = x.inscripcion.persona.direccion if x.inscripcion.persona.direccion else ""
                        campo19 = x.inscripcion.persona.direccion2 if x.inscripcion.persona.direccion2 else ""
                        campo20 = x.inscripcion.persona.sector if x.inscripcion.persona.sector else ""
                        campo21 = str(
                            x.matriculagruposocioeconomico().nombre) if x.matriculagruposocioeconomico() else ""
                        i += 1
                        ws.write(row_num, 0, i, font_style2)
                        ws.write(row_num, 1, campo1, font_style2)
                        ws.write(row_num, 2, campo2, font_style2)
                        ws.write(row_num, 3, campo3, font_style2)
                        ws.write(row_num, 4, campo4, font_style2)
                        ws.write(row_num, 5, campo5, font_style2)
                        ws.write(row_num, 6, campo6, font_style2)
                        ws.write(row_num, 7, campo7, font_style2)
                        ws.write(row_num, 8, campo8, font_style2)
                        ws.write(row_num, 9, campo9, font_style2)
                        ws.write(row_num, 10, campo11, font_style2)
                        ws.write(row_num, 11, campo12, font_style2)
                        ws.write(row_num, 12, campo13, font_style2)
                        ws.write(row_num, 13, campo14, font_style2)
                        ws.write(row_num, 14, campo15, font_style2)
                        ws.write(row_num, 15, campo10, font_style2)
                        ws.write(row_num, 16, campo16, font_style2)
                        ws.write(row_num, 17, campo17, font_style2)
                        ws.write(row_num, 18, campo18, font_style2)
                        ws.write(row_num, 19, campo19, font_style2)
                        ws.write(row_num, 20, campo20, font_style2)
                        ws.write(row_num, 21, campo21, font_style2)
                        ws.write(row_num, 22, str(periodoactual), font_style2)
                        ws.write(row_num, 23, str(periodovalida), font_style2)
                        row_num += 1
                wb.save(response)
                return response
            except Exception as ex:
                pass

        elif action == 'imprimir_promedio2':
            try:

                font_style = XFStyle()
                font_style.font.bold = True
                font_style2 = XFStyle()
                font_style2.font.bold = False
                wb = Workbook(encoding='utf-8')
                ws = wb.add_sheet('alumnos_promedio')
                response = HttpResponse(content_type="application/ms-excel")
                response['Content-Disposition'] = 'attachment; filename=alumnos' + random.randint(1,
                                                                                                  10000).__str__() + '.xls'
                columns = [
                    (u"N.", 1500),
                    (u"FACULTAD", 12000),
                    (u"CARRERA", 8000),
                    (u"CEDULA", 8000),
                    (u"PARALELO", 3000),
                    (u"PROMEDIO", 12000)
                ]
                row_num = 0
                for col_num in range(len(columns)):
                    ws.write(row_num, col_num, columns[col_num][0], font_style)
                    ws.col(col_num).width = columns[col_num][1]
                date_format = xlwt.XFStyle()
                date_format.num_format_str = 'yyyy/mm/dd'
                inscripcionesactuales = Matricula.objects.values_list('inscripcion__id', flat=True).filter(
                    status=True, nivel__periodo__id=periodo.id,
                    estado_matricula__in=[2, 3],
                    retiradomatricula=False,
                    matriculagruposocioeconomico__tipomatricula=1).exclude(inscripcion__persona__rubro__cancelado=False,
                                                                           inscripcion__persona__rubro__status=True,
                                                                           inscripcion__persona__rubro__fecha__lte=fechafinperiodonterior).distinct().order_by(
                    "inscripcion__carrera__coordinacion")
                nuevolistado = Matricula.objects.filter(status=True,
                                                        nivel__periodo__id=anterior.id,
                                                        estado_matricula__in=[2, 3],
                                                        retiradomatricula=False,
                                                        matriculagruposocioeconomico__tipomatricula=1,
                                                        inscripcion__id__in=inscripcionesactuales).exclude(
                    inscripcion__persona__rubro__cancelado=False,
                    inscripcion__persona__rubro__status=True,
                    inscripcion__persona__rubro__fecha__lte=fechafinperiodonterior).distinct().order_by(
                    "inscripcion__carrera__coordinacion")
                row_num = 1
                i = 0
                for x in nuevolistado:
                    materias = MateriaAsignada.objects.filter(status=True,
                                                              matricula__inscripcion__id=x.inscripcion.id,
                                                              matricula__nivel__periodo__id=anterior.id,
                                                              materiaasignadaretiro__isnull=True)
                    verifica = 0
                    suma = 0
                    promedio = 0
                    total = materias.count()
                    for m in materias:
                        suma += m.notafinal
                        if m.estado.id != 1:
                            verifica = 1
                            break
                    if suma > 0:
                        promedio = round(suma / total, 2)
                    if verifica == 0 and promedio >= 85:
                        cumple = requisito.validar_requisito_general(x.inscripcion.id, periodo.id)
                        if not cumple:
                            campo1 = x.inscripcion.coordinacion.nombre
                            if (x.inscripcion.carrera.mencion):
                                campo2 = x.inscripcion.carrera.nombre + ' CON MENCION EN  ' + x.inscripcion.carrera.mencion
                            else:
                                campo2 = x.inscripcion.carrera.nombre
                            campo3 = x.inscripcion.persona.cedula
                            if (x.paralelo):
                                campo4 = x.paralelo.nombre
                            else:
                                campo4 = 'Ninguno'
                            campo5 = promedio
                            i += 1
                            ws.write(row_num, 0, i, font_style2)
                            ws.write(row_num, 1, campo1, font_style2)
                            ws.write(row_num, 2, campo2, font_style2)
                            ws.write(row_num, 3, campo3, font_style2)
                            ws.write(row_num, 4, campo4, font_style2)
                            ws.write(row_num, 5, campo5, font_style2)
                            row_num += 1
                wb.save(response)
                return response
            except Exception as ex:
                pass

        elif action == 'imprimir_discapacitados':
            try:
                becatipo = BecaTipo.objects.get(pk=19)
                configuracion = becatipo.becatipoconfiguracion_set.filter(becaperiodo__periodo=periodo).first()

                action_extra = request.POST.get('a')
                cantidad_total_requisitos = 0
                nombre_archivo = 'alumnos_discapacitados'
                titulo_archivo = 'LISTADO DE ESTUDIANTE PRESELECCIONADO POR DISCAPACIDAD'
                if configuracion:
                    if action_extra:
                        cantidad_total_requisitos = configuracion.requisitosbecas.count() if configuracion.requisitosbecas else 0
                        if action_extra == 'requisitos_completos':
                            nombre_archivo = 'alumnos_discapacitados_OCAS'
                            titulo_archivo = 'LISTADO DE ESTUDIANTES SELECCIONADOS POR DISCAPACIDAD'
                        elif action_extra == 'requisitos_incompletos':
                            nombre_archivo = 'alumnos_discapacitados_rechazados'
                            titulo_archivo = 'LISTADO DE ESTUDIANTES NO SELECCIONADOS POR DISCAPACIDAD'


                font_style = XFStyle()
                font_style.font.bold = True
                font_style2 = XFStyle()
                font_style2.font.bold = False
                wb = Workbook(encoding='utf-8')
                # periodoactual = Periodo.objects.get(id=int(request.POST['periodoactual']))
                # periodovalida = Periodo.objects.get(id=int(request.POST['periodovalida']))
                periodoactual = periodosesion
                periodovalida = anterior
                # modalidad = int(request.POST['modalidad'])
                ws = wb.add_sheet('alumnos_discapacitados')
                response = HttpResponse(content_type="application/ms-excel")
                response['Content-Disposition'] = f'attachment; filename={nombre_archivo}' + random.randint(1, 10000).__str__() + '.xls'
                columns = [
                    (u"N.", 1500),
                    (u"ID_PREINSCRIPCION.", 1500),
                    (u"ORDEN", 1500),
                    (u"APELLIDOS Y NOMBRES", 12000),
                    (u"CARRERA", 8000),
                    (u"MODALIDAD", 8000),
                    (u"CEDULA", 3000),
                    (u"FECHA NACIMIENTO", 3000),
                    (u"DISCAPACIDAD", 3000),
                    (u"PORCENTAJE", 3000),
                    (u"CARNET", 3000),
                    (u"PROMEDIO", 3000),
                    (u"ASISTENCIA", 3000),
                    (u"SESION", 3000),
                    (u"NIVEL", 3000),
                    (u"PARALELO", 3000),
                    (u"DIRECCION", 3000),
                    (u"TELEFONO", 4000),
                    (u"TELEFONO CONVENCIONAL", 4000),
                    (u"EMAIL", 5000),
                    (u"EMAIL INSTITUCIONAL", 5000),
                    (u"SEXO", 5000),
                    (u"PAIS", 5000),
                    (u"PROVINCIA", 5000),
                    (u"CANTON", 5000),
                    (u"DIRECCION 1", 5000),
                    (u"DIRECCION 2", 5000),
                    (u"SECTOR", 5000),
                    (u"GRUPO SOCIOECONOMICO", 5000),
                    (u"PERIODO ACTUAL", 8000),
                    (u"PERIDOO VALIDA", 8000),
                    (u"TIPO/ESTADO", 2500),
                ]

                requisitos = []
                if configuracion:
                    #Requisitos de tipos de becas
                    requisitos = configuracion.requisitosbecas.filter(visible=True, status=True)
                    for detallerequisitobeca in requisitos:
                        columns.append((f'({detallerequisitobeca.pk}) {detallerequisitobeca.requisitobeca.__str__()}', 10000))
                        if not action_extra:
                            columns.append((f'({detallerequisitobeca.pk}) OBSERVACIÓN', 5000))

                style_title = xlwt.easyxf(
                    'font: height 350, name Arial, colour_index black, bold on, italic on; align: wrap on, vert centre, horiz center;')
                style_title_2 = xlwt.easyxf(
                    'font: height 350, name Arial, colour_index black, bold on, italic on; align: wrap on, vert centre, horiz center;')
                ws.write_merge(0, 0, 0, len(columns), 'UNIVERSIDAD ESTATAL DE MILAGRO', style_title)
                ws.write_merge(1, 1, 0, len(columns), titulo_archivo, style_title_2)
                row_num = 3
                for col_num in range(len(columns)):
                    ws.write(row_num, col_num, columns[col_num][0], font_style)
                    ws.col(col_num).width = columns[col_num][1]
                date_format = xlwt.XFStyle()
                date_format.num_format_str = 'yyyy/mm/dd'
                if periodosesion.versionbeca is None or periodosesion.versionbeca == 1:
                    discapacitados = Matricula.objects.values_list('inscripcion__id', flat=True).filter(status=True,
                                                                                                        nivel__periodo=periodoactual,
                                                                                                        estado_matricula__in=[
                                                                                                            2, 3],
                                                                                                        retiradomatricula=False,
                                                                                                        inscripcion__persona__perfilinscripcion__tienediscapacidad=True,
                                                                                                        inscripcion__persona__perfilinscripcion__verificadiscapacidad=True,
                                                                                                        matriculagruposocioeconomico__tipomatricula=1).exclude(
                        inscripcion__carrera__coordinacion__id__in=[7, 9]).distinct().order_by("inscripcion__persona")
                    matriculados = Matricula.objects.filter(status=True, nivel__periodo=periodovalida,
                                                            estado_matricula__in=[2, 3],
                                                            retiradomatricula=False,
                                                            matriculagruposocioeconomico__tipomatricula=1,
                                                            inscripcion__id__in=discapacitados).exclude(
                        inscripcion__carrera__coordinacion__id__in=[7, 9]).distinct().order_by("inscripcion__persona")
                elif periodosesion.versionbeca == 2:
                    inscripciones = PreInscripcionBeca.objects.filter(periodo=periodosesion, becatipo_id=19)

                    if action_extra == 'requisitos_completos':
                        inscripciones = inscripciones.filter(preinscripcionbecarequisito__cumplerequisito=True)\
                            .annotate(total_requisitos=Count('preinscripcionbecarequisito', filter=Q(status=True), distinct=True))\
                            .filter(total_requisitos=cantidad_total_requisitos)
                    elif action_extra == 'requisitos_incompletos':
                        pendientes = inscripciones.values_list('id', flat=True).filter(preinscripcionbecarequisito__cumplerequisito__isnull=True)
                        inscripciones = inscripciones.filter(preinscripcionbecarequisito__cumplerequisito=False).exclude(id__in=pendientes).distinct()

                    matriculados = Matricula.objects.filter(inscripcion__id__in=inscripciones.values_list('inscripcion_id',flat=True),
                                                            nivel__periodo__id=periodo.id).distinct().order_by(
                        "inscripcion__persona")
                row_num = 4
                i = 0
                for x in matriculados:
                    preinscripcion = PreInscripcionBeca.objects.filter(inscripcion=x.inscripcion, periodo=periodo, becatipo_id=19).first()
                    x1 = InscripcionNivel.objects.filter(inscripcion=x.inscripcion, nivel__orden=1)
                    isPrimerNivel = False
                    verifica = 0
                    suma = 0
                    promedio = 0
                    sumasis = 0
                    asistencia = 0
                    if x1.exists():
                        isPrimerNivel = True
                    if not isPrimerNivel:
                        asignaturas = MateriaAsignada.objects.filter(status=True,
                                                                     retiramateria=False,
                                                                     matricula__inscripcion__id=x.inscripcion.id,
                                                                     matricula__nivel__periodo__id=anterior.id,
                                                                     materiaasignadaretiro__isnull=True).exclude(materia__asignaturamalla__malla_id__in=[353, 22])

                        total = asignaturas.count()
                        for m in asignaturas:
                            suma += m.notafinal
                            sumasis += m.asistenciafinal
                            if m.estado.id != 1:
                                verifica = 1
                                break
                        if suma > 0 and verifica == 0:
                            promedio = round(suma / total, 2)
                            asistencia = round(sumasis / total, 2)
                    #if verifica == 0:
                    campo1 = preinscripcion.id
                    campo2 = preinscripcion.orden
                    campo3 = x.inscripcion.persona.nombre_completo_inverso()
                    if x.inscripcion.carrera.mencion:
                        campo4 = x.inscripcion.carrera.nombre + ' CON MENCION EN  ' + x.inscripcion.carrera.mencion
                    else:
                        campo4 = x.inscripcion.carrera.nombre
                    campo5 = x.inscripcion.modalidad.__str__()
                    campo6 = x.inscripcion.persona.cedula
                    campo7 = x.inscripcion.persona.nacimiento
                    if periodo.id >= 119:
                        campo8 = preinscripcion.tipodiscapacidad.nombre if preinscripcion.tipodiscapacidad else ""
                        campo9 = preinscripcion.porcientodiscapacidad
                        campo10 = preinscripcion.carnetdiscapacidad
                    else:
                        campo8 = x.inscripcion.persona.mi_perfil().tipodiscapacidad.nombre if x.inscripcion.persona.mi_perfil().tipodiscapacidad else ""
                        campo9 = x.inscripcion.persona.mi_perfil().porcientodiscapacidad
                        campo10 = x.inscripcion.persona.mi_perfil().carnetdiscapacidad
                    campo11 = promedio if not isPrimerNivel and promedio > 0 else 'NO APLICA'
                    campo12 = asistencia if not isPrimerNivel and asistencia > 0 else 'NO APLICA'
                    campo13 = x.inscripcion.sesion.nombre
                    if x.nivelmalla:
                        campo14 = x.nivelmalla.nombre
                    else:
                        campo14 = 'Ninguno'
                    if x.paralelo:
                        campo15 = x.paralelo.nombre
                    else:
                        campo15 = 'Ninguno'
                    campo16 = x.inscripcion.persona.direccion_completa()
                    campo17 = x.inscripcion.persona.telefono if x.inscripcion.persona.telefono else ""
                    campo18 = x.inscripcion.persona.telefono_conv if x.inscripcion.persona.telefono_conv else ""
                    campo19 = x.inscripcion.persona.email if x.inscripcion.persona.email else ""
                    campo20 = x.inscripcion.persona.emailinst if x.inscripcion.persona.emailinst else ""
                    campo21 = x.inscripcion.persona.sexo.nombre if x.inscripcion.persona.sexo else ""
                    campo22 = x.inscripcion.persona.pais.nombre if x.inscripcion.persona.pais else ""
                    campo23 = x.inscripcion.persona.provincia.nombre if x.inscripcion.persona.provincia else ""
                    campo24 = x.inscripcion.persona.canton.nombre if x.inscripcion.persona.canton else ""
                    campo25 = x.inscripcion.persona.direccion if x.inscripcion.persona.direccion else ""
                    campo26 = x.inscripcion.persona.direccion2 if x.inscripcion.persona.direccion2 else ""
                    campo27 = x.inscripcion.persona.sector if x.inscripcion.persona.sector else ""
                    campo28 = str(
                        x.matriculagruposocioeconomico().nombre) if x.matriculagruposocioeconomico() else ""
                    campo29 = x.estado_renovacion_beca(becatipo, anterior)
                    i += 1
                    ws.write(row_num, 0, i, font_style2)
                    ws.write(row_num, 1, campo1, font_style2)
                    ws.write(row_num, 2, campo2, font_style2)
                    ws.write(row_num, 3, campo3, font_style2)
                    ws.write(row_num, 4, campo4, font_style2)
                    ws.write(row_num, 5, campo5, font_style2)
                    ws.write(row_num, 6, campo6, font_style2)
                    ws.write(row_num, 7, campo7, date_format)
                    ws.write(row_num, 8, campo8, font_style2)
                    ws.write(row_num, 9, campo9, font_style2)
                    ws.write(row_num, 10, campo10, font_style2)
                    ws.write(row_num, 11, campo11, font_style2)
                    ws.write(row_num, 12, campo12, font_style2)
                    ws.write(row_num, 13, campo13, font_style2)
                    ws.write(row_num, 14, campo14, font_style2)
                    ws.write(row_num, 15, campo15, font_style2)
                    ws.write(row_num, 16, campo16, font_style2)
                    ws.write(row_num, 17, campo17, font_style2)
                    ws.write(row_num, 18, campo18, font_style2)
                    ws.write(row_num, 19, campo19, font_style2)
                    ws.write(row_num, 20, campo20, font_style2)
                    ws.write(row_num, 21, campo21, font_style2)
                    ws.write(row_num, 22, campo22, font_style2)
                    ws.write(row_num, 23, campo23, font_style2)
                    ws.write(row_num, 24, campo24, font_style2)
                    ws.write(row_num, 25, campo25, font_style2)
                    ws.write(row_num, 26, campo26, font_style2)
                    ws.write(row_num, 27, campo27, font_style2)
                    ws.write(row_num, 28, campo28, font_style2)
                    ws.write(row_num, 29, str(periodoactual), font_style2)
                    ws.write(row_num, 30, str(periodovalida), font_style2)
                    ws.write(row_num, 31, campo29, font_style2)

                    if requisitos:
                        col_num = len(columns) - requisitos.count()*(2 if not action_extra else 1)
                        for detallerequisitobeca in requisitos:
                            preinscriocionrequisitos = PreInscripcionBecaRequisito.objects.filter(preinscripcion=preinscripcion, detallerequisitobeca=detallerequisitobeca, status=True)
                            for preinsrequisito in preinscriocionrequisitos:
                                cumple = ''
                                cumplerequisito = preinsrequisito.cumplerequisito
                                if cumplerequisito:
                                   cumple = 'SI'
                                elif cumplerequisito == False:
                                   cumple = 'NO'
                                ws.write(row_num, col_num, cumple, font_style2)

                                if not action_extra:
                                    col_num += 1
                                    ws.write(row_num, col_num, '', font_style2) #Observación por columna
                            col_num += 1
                    row_num += 1
                wb.save(response)
                return response
            except Exception as ex:
                err = 'Error on line {}'.format(sys.exc_info()[-1].tb_lineno)
                msg = ex.__str__()
                msg = f'{msg} {err}'

                return JsonResponse({"result": "bad", "mensaje": u"Error: %s"%msg})

        elif action == 'imprimir_etnias':
            try:
                becatipo = BecaTipo.objects.get(pk=21)
                configuracion = becatipo.becatipoconfiguracion_set.filter(becaperiodo__periodo=periodo).first()

                action_extra = request.POST.get('a')
                cantidad_total_requisitos = 0
                nombre_archivo = 'alumnos_etnias'
                titulo_archivo = 'LISTADO DE ESTUDIANTE PRESELECCIONADO POR PERTENCER A PUEBLOS Y NACIONALIDADES DEL ECUADOR'
                if configuracion:
                    if action_extra:
                        cantidad_total_requisitos = configuracion.requisitosbecas.count() if configuracion.requisitosbecas else 0
                        if action_extra == 'requisitos_completos':
                            nombre_archivo = 'alumnos_etnias_OCAS'
                            titulo_archivo = 'LISTADO DE ESTUDIANTES SELECCIONADOS POR PERTENCER A PUEBLOS Y NACIONALIDADES DEL ECUADOR'
                        elif action_extra == 'requisitos_incompletos':
                            nombre_archivo = 'alumnos_etnias_rechazados'
                            titulo_archivo = 'LISTADO DE ESTUDIANTES NO SELECCIONADOS POR PERTENCER A PUEBLOS Y NACIONALIDADES DEL ECUADOR'


                font_style = XFStyle()
                font_style.font.bold = True
                font_style2 = XFStyle()
                font_style2.font.bold = False
                wb = Workbook(encoding='utf-8')
                # periodoactual = Periodo.objects.get(id=int(request.POST['periodoactual']))
                # periodovalida = Periodo.objects.get(id=int(request.POST['periodovalida']))
                periodoactual = periodosesion
                periodovalida = anterior
                # modalidad = int(request.POST['modalidad'])
                ws = wb.add_sheet('alumnos_etnias')
                response = HttpResponse(content_type="application/ms-excel")
                response['Content-Disposition'] = f'attachment; filename={nombre_archivo}' + random.randint(1,
                                                                                                         10000).__str__() + '.xls'
                columns = [
                    (u"N.", 1500),
                    (u"ID_PREINSCRIPCION.", 1500),
                    (u"ORDEN.", 1000),
                    (u"APELLIDOS Y NOMBRES", 12000),
                    (u"CARRERA", 8000),
                    (u"MODALIDAD", 8000),
                    (u"CEDULA", 3000),
                    (u"FECHA NACIMIENTO", 3000),
                    (u"ETNIA", 3000),
                    (u"PROMEDIO", 3000),
                    (u"ASISTENCIA", 3000),
                    (u"SESION", 3000),
                    (u"NIVEL", 3000),
                    (u"PARALELO", 3000),
                    (u"DIRECCION", 3000),
                    (u"TELEFONO", 4000),
                    (u"TELEFONO CONVENCIONAL", 4000),
                    (u"EMAIL", 5000),
                    (u"EMAIL INSTITUCIONAL", 5000),
                    (u"SEXO", 5000),
                    (u"PAIS", 5000),
                    (u"PROVINCIA", 5000),
                    (u"CANTON", 5000),
                    (u"DIRECCION 1", 5000),
                    (u"DIRECCION 2", 5000),
                    (u"SECTOR", 5000),
                    (u"GRUPO SOCIOECONOMICO", 5000),
                    (u"PERIODO ACTUAL", 8000),
                    (u"PERIDOO VALIDA", 8000),
                    (u"TIPO/ESTADO", 2500),
                ]

                requisitos = []
                if configuracion:
                    #Requisitos de tipos de becas
                    requisitos = configuracion.requisitosbecas.filter(visible=True, status=True)
                    for detallerequisitobeca in requisitos:
                        columns.append((f'({detallerequisitobeca.pk}) {detallerequisitobeca.requisitobeca.__str__()}', 10000))
                        if not action_extra:
                            columns.append((f'({detallerequisitobeca.pk}) OBSERVACIÓN', 5000))

                style_title = xlwt.easyxf(
                    'font: height 350, name Arial, colour_index black, bold on, italic on; align: wrap on, vert centre, horiz center;')
                style_title_2 = xlwt.easyxf(
                    'font: height 350, name Arial, colour_index black, bold on, italic on; align: wrap on, vert centre, horiz center;')
                ws.write_merge(0, 0, 0, len(columns), 'UNIVERSIDAD ESTATAL DE MILAGRO', style_title_2)
                ws.write_merge(1, 1, 0, len(columns), titulo_archivo, font_style)
                row_num = 3
                for col_num in range(len(columns)):
                    ws.write(row_num, col_num, columns[col_num][0], font_style)
                    ws.col(col_num).width = columns[col_num][1]
                date_format = xlwt.XFStyle()
                date_format.num_format_str = 'yyyy/mm/dd'
                if periodosesion.versionbeca is None or periodosesion.versionbeca == 1:
                    discapacitados = Matricula.objects.values_list('inscripcion__id', flat=True).filter(status=True,
                                                                                                        nivel__periodo=periodoactual,
                                                                                                        estado_matricula__in=[
                                                                                                            2, 3],
                                                                                                        retiradomatricula=False,
                                                                                                        inscripcion__persona__perfilinscripcion__tienediscapacidad=True,
                                                                                                        inscripcion__persona__perfilinscripcion__verificadiscapacidad=True,
                                                                                                        matriculagruposocioeconomico__tipomatricula=1).exclude(
                        inscripcion__carrera__coordinacion__id__in=[7, 9]).distinct().order_by("inscripcion__persona")
                    matriculados = Matricula.objects.filter(status=True, nivel__periodo=periodovalida,
                                                            estado_matricula__in=[2, 3],
                                                            retiradomatricula=False,
                                                            matriculagruposocioeconomico__tipomatricula=1,
                                                            inscripcion__id__in=discapacitados).exclude(
                        inscripcion__carrera__coordinacion__id__in=[7, 9]).distinct().order_by("inscripcion__persona")
                elif periodosesion.versionbeca == 2:
                    inscripciones = PreInscripcionBeca.objects.filter(periodo=periodosesion,
                                                                      becatipo_id=21).values_list('inscripcion_id',
                                                                                                  flat=True)
                    if action_extra == 'requisitos_completos':
                        inscripciones = inscripciones.filter(preinscripcionbecarequisito__cumplerequisito=True)\
                            .annotate(total_requisitos=Count('preinscripcionbecarequisito', filter=Q(status=True), distinct=True))\
                            .filter(total_requisitos=cantidad_total_requisitos)
                    elif action_extra == 'requisitos_incompletos':
                        pendientes = inscripciones.values_list('id', flat=True).filter(preinscripcionbecarequisito__cumplerequisito__isnull=True)
                        inscripciones = inscripciones.filter(preinscripcionbecarequisito__cumplerequisito=False).exclude(id__in=pendientes).distinct()

                    matriculados = Matricula.objects.filter(inscripcion__id__in=inscripciones.values_list('inscripcion_id',flat=True),
                                                            nivel__periodo__id=periodo.id).distinct().order_by(
                        "inscripcion__persona")
                row_num = 4
                i = 0
                for x in matriculados:
                    preinscripcion = PreInscripcionBeca.objects.filter(inscripcion=x.inscripcion, periodo=periodo, becatipo_id=21).first()
                    asignaturas = MateriaAsignada.objects.filter(status=True,
                                                                 retiramateria=False,
                                                                 matricula__inscripcion__id=x.inscripcion.id,
                                                                 matricula__nivel__periodo__id=anterior.id,
                                                                 materiaasignadaretiro__isnull=True).exclude(materia__asignaturamalla__malla_id__in=[353, 22])
                    verifica = 0
                    suma = 0
                    promedio = 0
                    sumasis = 0
                    total = asignaturas.count()
                    for m in asignaturas:
                        suma += m.notafinal
                        sumasis += m.asistenciafinal
                        if m.estado.id != 1:
                            verifica = 1
                            break
                    if suma > 0 and verifica == 0:
                        promedio = round(suma / total, 2)
                        asistencia = round(sumasis / total, 2)

                    #if verifica == 0:
                    campo1 = preinscripcion.id
                    campo2 = preinscripcion.orden
                    campo3 = x.inscripcion.persona.nombre_completo_inverso()
                    if x.inscripcion.carrera.mencion:
                        campo4 = x.inscripcion.carrera.nombre + ' CON MENCION EN  ' + x.inscripcion.carrera.mencion
                    else:
                        campo4 = x.inscripcion.carrera.nombre
                    campo5 = x.inscripcion.modalidad.__str__()
                    campo6 = x.inscripcion.persona.cedula
                    campo7 = x.inscripcion.persona.nacimiento
                    if periodo.id >= 119:
                        campo8 = preinscripcion.raza.nombre if preinscripcion.raza else '' #x.inscripcion.persona.mi_perfil().raza.nombre
                    else:
                        campo8 = x.inscripcion.persona.mi_perfil().raza.nombre
                    campo9 = promedio
                    campo10 = asistencia
                    campo11 = x.inscripcion.sesion.nombre
                    if x.nivelmalla:
                        campo12 = x.nivelmalla.nombre
                    else:
                        campo12 = 'Ninguno'
                    if x.paralelo:
                        campo13 = x.paralelo.nombre
                    else:
                        campo13 = 'Ninguno'
                    campo14 = x.inscripcion.persona.direccion_completa()
                    campo15 = x.inscripcion.persona.telefono
                    campo16 = x.inscripcion.persona.telefono_conv
                    campo17 = x.inscripcion.persona.email
                    campo18 = x.inscripcion.persona.emailinst
                    campo19 = x.inscripcion.persona.sexo.nombre
                    campo20 = x.inscripcion.persona.pais.nombre if x.inscripcion.persona.pais else ""
                    campo21 = x.inscripcion.persona.provincia.nombre if x.inscripcion.persona.provincia else ""
                    campo22 = x.inscripcion.persona.canton.nombre if x.inscripcion.persona.canton else ""
                    campo23 = x.inscripcion.persona.direccion if x.inscripcion.persona.direccion else ""
                    campo24 = x.inscripcion.persona.direccion2 if x.inscripcion.persona.direccion2 else ""
                    campo25 = x.inscripcion.persona.sector if x.inscripcion.persona.sector else ""
                    campo26 = str(
                        x.matriculagruposocioeconomico().nombre) if x.matriculagruposocioeconomico() else ""
                    campo27 = x.estado_renovacion_beca(becatipo, anterior)
                    i += 1
                    ws.write(row_num, 0, i, font_style2)
                    ws.write(row_num, 1, campo1, font_style2)
                    ws.write(row_num, 2, campo2, font_style2)
                    ws.write(row_num, 3, campo3, font_style2)
                    ws.write(row_num, 4, campo4, font_style2)
                    ws.write(row_num, 5, campo5, font_style2)
                    ws.write(row_num, 6, campo6, font_style2)
                    ws.write(row_num, 7, campo7, date_format)
                    ws.write(row_num, 8, campo8, font_style2)
                    ws.write(row_num, 9, campo9, font_style2)
                    ws.write(row_num, 10, campo10, font_style2)
                    ws.write(row_num, 11, campo11, font_style2)
                    ws.write(row_num, 12, campo12, font_style2)
                    ws.write(row_num, 13, campo13, font_style2)
                    ws.write(row_num, 14, campo14, font_style2)
                    ws.write(row_num, 15, campo15, font_style2)
                    ws.write(row_num, 16, campo16, font_style2)
                    ws.write(row_num, 17, campo17, font_style2)
                    ws.write(row_num, 18, campo18, font_style2)
                    ws.write(row_num, 19, campo19, font_style2)
                    ws.write(row_num, 20, campo20, font_style2)
                    ws.write(row_num, 21, campo21, font_style2)
                    ws.write(row_num, 22, campo22, font_style2)
                    ws.write(row_num, 23, campo23, font_style2)
                    ws.write(row_num, 24, campo24, font_style2)
                    ws.write(row_num, 25, campo25, font_style2)
                    ws.write(row_num, 26, campo26, font_style2)
                    ws.write(row_num, 27, str(periodoactual), font_style2)
                    ws.write(row_num, 28, str(periodovalida), font_style2)
                    ws.write(row_num, 29, campo27, font_style2)

                    if requisitos:
                        col_num = len(columns) - requisitos.count()*(2 if not action_extra else 1)
                        for detallerequisitobeca in requisitos:
                            preinscriocionrequisitos = PreInscripcionBecaRequisito.objects.filter(preinscripcion=preinscripcion, detallerequisitobeca=detallerequisitobeca, status=True)
                            for preinsrequisito in preinscriocionrequisitos:
                                cumple = ''
                                cumplerequisito = preinsrequisito.cumplerequisito
                                if cumplerequisito:
                                   cumple = 'SI'
                                elif cumplerequisito == False:
                                   cumple = 'NO'
                                ws.write(row_num, col_num, cumple, font_style2)

                                if not action_extra:
                                    col_num += 1
                                    ws.write(row_num, col_num, '', font_style2) #Observación por columna
                            col_num += 1
                    row_num += 1
                wb.save(response)
                return response
            except Exception as ex:
                err = 'Error on line {}'.format(sys.exc_info()[-1].tb_lineno)
                msg = ex.__str__()
                msg = f'{msg} {err}'

                return JsonResponse({"result": "bad", "mensaje": u"Error: %s"%msg})

        elif action == 'imprimir_extranejeros':
            try:
                becatipo = BecaTipo.objects.get(pk=22)
                configuracion = becatipo.becatipoconfiguracion_set.filter(becaperiodo__periodo=periodo).first()

                action_extra = request.POST.get('a')
                cantidad_total_requisitos = 0
                nombre_archivo = 'alumnos_extranjeros'
                titulo_archivo = 'LISTADO DE ESTUDIANTE PRESELECCIONADO POR SER ECUATORIANO EN EL EXTERIOR'
                if configuracion:
                    if action_extra:
                        cantidad_total_requisitos = configuracion.requisitosbecas.count() if configuracion.requisitosbecas else 0
                        if action_extra == 'requisitos_completos':
                            nombre_archivo = 'alumnos_extranjeros_OCAS'
                            titulo_archivo = 'LISTADO DE ESTUDIANTES SELECCIONADOS POR SER ECUATORIANO EN EL EXTERIOR'
                        elif action_extra == 'requisitos_incompletos':
                            nombre_archivo = 'alumnos_extranjeros_rechazados'
                            titulo_archivo = 'LISTADO DE ESTUDIANTES NO SELECCIONADOS POR SER ECUATORIANO EN EL EXTERIOR'

                font_style = XFStyle()
                font_style.font.bold = True
                font_style2 = XFStyle()
                font_style2.font.bold = False
                wb = Workbook(encoding='utf-8')
                # periodoactual = Periodo.objects.get(id=int(request.POST['periodoactual']))
                # periodovalida = Periodo.objects.get(id=int(request.POST['periodovalida']))
                periodoactual = periodosesion
                periodovalida = anterior
                # modalidad = int(request.POST['modalidad'])
                ws = wb.add_sheet('alumnos_extranjeros')
                response = HttpResponse(content_type="application/ms-excel")
                response['Content-Disposition'] =  f'attachment; filename={nombre_archivo}' + random.randint(1,
                                                                                                              10000).__str__() + '.xls'
                columns = [
                    (u"N.", 1500),
                    (u"ID_PREINSCRIPCION.", 1500),
                    (u"ORDEN", 1500),
                    (u"APELLIDOS Y NOMBRES", 12000),
                    (u"CARRERA", 8000),
                    (u"MODALIDAD", 8000),
                    (u"CEDULA", 3000),
                    (u"FECHA NACIMIENTO", 3000),
                    (u"PAIS EXTERIOR", 3000),
                    (u"PROMEDIO", 3000),
                    (u"ASISTENCIA", 3000),
                    (u"SESION", 3000),
                    (u"NIVEL", 3000),
                    (u"PARALELO", 3000),
                    (u"DIRECCION", 3000),
                    (u"TELEFONO", 4000),
                    (u"TELEFONO CONVENCIONAL", 4000),
                    (u"EMAIL", 5000),
                    (u"EMAIL INSTITUCIONAL", 5000),
                    (u"SEXO", 5000),
                    (u"PAIS", 5000),
                    (u"PROVINCIA", 5000),
                    (u"CANTON", 5000),
                    (u"DIRECCION 1", 5000),
                    (u"DIRECCION 2", 5000),
                    (u"SECTOR", 5000),
                    (u"GRUPO SOCIOECONOMICO", 5000),
                    (u"PERIODO ACTUAL", 8000),
                    (u"PERIDOO VALIDA", 8000),
                    (u"TIPO/ESTADO", 2500),
                ]

                requisitos = []
                if configuracion:
                    #Requisitos de tipos de becas
                    requisitos = configuracion.requisitosbecas.filter(visible=True, status=True)
                    for detallerequisitobeca in requisitos:
                        columns.append((f'({detallerequisitobeca.pk}) {detallerequisitobeca.requisitobeca.__str__()}', 10000))
                        if not action_extra:
                            columns.append((f'({detallerequisitobeca.pk}) OBSERVACIÓN', 5000))

                style_title = xlwt.easyxf(
                    'font: height 350, name Arial, colour_index black, bold on, italic on; align: wrap on, vert centre, horiz center;')
                style_title_2 = xlwt.easyxf(
                    'font: height 350, name Arial, colour_index black, bold on, italic on; align: wrap on, vert centre, horiz center;')
                ws.write_merge(0, 0, 0, len(columns), 'UNIVERSIDAD ESTATAL DE MILAGRO', style_title)
                ws.write_merge(1, 1, 0, len(columns), titulo_archivo, style_title_2)
                row_num = 3
                for col_num in range(len(columns)):
                    ws.write(row_num, col_num, columns[col_num][0], font_style)
                    ws.col(col_num).width = columns[col_num][1]
                date_format = xlwt.XFStyle()
                date_format.num_format_str = 'yyyy/mm/dd'
                if periodosesion.versionbeca is None or periodosesion.versionbeca == 1:
                    matriculados = Matricula.objects.filter(pk=None)
                elif periodosesion.versionbeca == 2:
                    inscripciones = PreInscripcionBeca.objects.filter(periodo=periodosesion,
                                                                      becatipo_id=22).values_list('inscripcion_id',
                                                                                                  flat=True)

                    if action_extra == 'requisitos_completos':
                        inscripciones = inscripciones.filter(preinscripcionbecarequisito__cumplerequisito=True)\
                            .annotate(total_requisitos=Count('preinscripcionbecarequisito', filter=Q(status=True), distinct=True))\
                            .filter(total_requisitos=cantidad_total_requisitos)
                    elif action_extra == 'requisitos_incompletos':
                        pendientes = inscripciones.values_list('id', flat=True).filter(preinscripcionbecarequisito__cumplerequisito__isnull=True)
                        inscripciones = inscripciones.filter(preinscripcionbecarequisito__cumplerequisito=False).exclude(id__in=pendientes).distinct()

                    matriculados = Matricula.objects.filter(~Q(inscripcion__persona__pais_id=1),
                                                            inscripcion__id__in=inscripciones.values_list('inscripcion_id',flat=True),
                                                            nivel__periodo__id=periodo.id).distinct().order_by(
                        "inscripcion__persona")
                row_num = 4
                i = 0
                becatipo = BecaTipo.objects.get(pk=22)
                for x in matriculados:
                    preinscripcion = PreInscripcionBeca.objects.filter(inscripcion=x.inscripcion, periodo=periodo, becatipo_id=22).first()
                    asignaturas = MateriaAsignada.objects.filter(status=True,
                                                                 retiramateria=False,
                                                                 matricula__inscripcion__id=x.inscripcion.id,
                                                                 matricula__nivel__periodo__id=anterior.id,
                                                                 materiaasignadaretiro__isnull=True).exclude(
                            materia__asignaturamalla__malla_id__in=[353, 22])
                    verifica = 0
                    suma = 0
                    promedio = 0
                    sumasis = 0
                    asistencia = 100
                    total = asignaturas.count()
                    for m in asignaturas:
                        suma += m.notafinal
                        sumasis += m.asistenciafinal
                        if m.estado.id != 1:
                            verifica = 1
                            break
                    if suma > 0 and verifica == 0:
                        promedio = round(suma / total, 2)
                        asistencia = round(sumasis / total, 2)
                    #if verifica == 0:
                    campo1 = preinscripcion.id
                    campo2 = preinscripcion.orden
                    campo3 = x.inscripcion.persona.nombre_completo_inverso()
                    if x.inscripcion.carrera.mencion:
                        campo4 = x.inscripcion.carrera.nombre + ' CON MENCION EN  ' + x.inscripcion.carrera.mencion
                    else:
                        campo4 = x.inscripcion.carrera.nombre
                    campo5 = x.inscripcion.modalidad.__str__()
                    campo6 = x.inscripcion.persona.cedula
                    campo7 = x.inscripcion.persona.nacimiento
                    campo8 = x.inscripcion.persona.pais.nombre
                    campo9 = promedio
                    campo10 = asistencia
                    campo11 = x.inscripcion.sesion.nombre
                    if x.nivelmalla:
                        campo12 = x.nivelmalla.nombre
                    else:
                        campo12 = 'Ninguno'
                    if x.paralelo:
                        campo13 = x.paralelo.nombre
                    else:
                        campo13 = 'Ninguno'
                    campo14 = x.inscripcion.persona.direccion_completa()
                    campo15 = x.inscripcion.persona.telefono
                    campo16 = x.inscripcion.persona.telefono_conv
                    campo17 = x.inscripcion.persona.email
                    campo18 = x.inscripcion.persona.emailinst
                    campo19 = x.inscripcion.persona.sexo.nombre
                    campo20 = x.inscripcion.persona.pais.nombre if x.inscripcion.persona.pais else ""
                    campo21 = x.inscripcion.persona.provincia.nombre if x.inscripcion.persona.provincia else ""
                    campo22 = x.inscripcion.persona.canton.nombre if x.inscripcion.persona.canton else ""
                    campo23 = x.inscripcion.persona.direccion if x.inscripcion.persona.direccion else ""
                    campo24 = x.inscripcion.persona.direccion2 if x.inscripcion.persona.direccion2 else ""
                    campo25 = x.inscripcion.persona.sector if x.inscripcion.persona.sector else ""
                    campo26 = str(
                        x.matriculagruposocioeconomico().nombre) if x.matriculagruposocioeconomico() else ""
                    campo27 = x.estado_renovacion_beca(becatipo, anterior)
                    i += 1
                    ws.write(row_num, 0, i, font_style2)
                    ws.write(row_num, 1, campo1, font_style2)
                    ws.write(row_num, 2, campo2, font_style2)
                    ws.write(row_num, 3, campo3, font_style2)
                    ws.write(row_num, 4, campo4, font_style2)
                    ws.write(row_num, 5, campo5, font_style2)
                    ws.write(row_num, 6, campo6, font_style2)
                    ws.write(row_num, 7, campo7, date_format)
                    ws.write(row_num, 8, campo8, font_style2)
                    ws.write(row_num, 9, campo9, font_style2)
                    ws.write(row_num, 10, campo10, font_style2)
                    ws.write(row_num, 11, campo11, font_style2)
                    ws.write(row_num, 12, campo12, font_style2)
                    ws.write(row_num, 13, campo13, font_style2)
                    ws.write(row_num, 14, campo14, font_style2)
                    ws.write(row_num, 15, campo15, font_style2)
                    ws.write(row_num, 16, campo16, font_style2)
                    ws.write(row_num, 17, campo17, font_style2)
                    ws.write(row_num, 18, campo18, font_style2)
                    ws.write(row_num, 19, campo19, font_style2)
                    ws.write(row_num, 20, campo20, font_style2)
                    ws.write(row_num, 21, campo21, font_style2)
                    ws.write(row_num, 22, campo22, font_style2)
                    ws.write(row_num, 23, campo23, font_style2)
                    ws.write(row_num, 24, campo24, font_style2)
                    ws.write(row_num, 25, campo25, font_style2)
                    ws.write(row_num, 26, campo26, font_style2)
                    ws.write(row_num, 27, str(periodoactual), font_style2)
                    ws.write(row_num, 28, str(periodovalida), font_style2)
                    ws.write(row_num, 29, campo27, font_style2)

                    if requisitos:
                        col_num = len(columns) - requisitos.count()*(2 if not action_extra else 1)
                        for detallerequisitobeca in requisitos:
                            preinscriocionrequisitos = PreInscripcionBecaRequisito.objects.filter(preinscripcion=preinscripcion, detallerequisitobeca=detallerequisitobeca, status=True)
                            for preinsrequisito in preinscriocionrequisitos:
                                cumple = ''
                                cumplerequisito = preinsrequisito.cumplerequisito
                                if cumplerequisito:
                                   cumple = 'SI'
                                elif cumplerequisito == False:
                                   cumple = 'NO'
                                ws.write(row_num, col_num, cumple, font_style2)
                                if not action_extra:
                                    col_num += 1
                                    ws.write(row_num, col_num, '', font_style2) #Observación por columna
                            col_num += 1
                    row_num += 1
                wb.save(response)
                return response
            except Exception as ex:
                err = 'Error on line {}'.format(sys.exc_info()[-1].tb_lineno)
                msg = ex.__str__()
                msg = f'{msg} {err}'

                return JsonResponse({"result": "bad", "mensaje": u"Error: %s"%msg})

        elif action == 'imprimir_migrantes':
            try:
                becatipo = BecaTipo.objects.get(pk=22)
                configuracion = becatipo.becatipoconfiguracion_set.filter(becaperiodo__periodo=periodo).first()

                action_extra = request.POST.get('a')
                cantidad_total_requisitos = 0
                nombre_archivo = 'alumnos_migrantes'
                titulo_archivo = 'LISTADO DE ESTUDIANTE PRESELECCIONADO POR SER ECUATORIANO MIGRANTE RETORNADO O DEPORTADO'
                if configuracion:
                    if action_extra:
                        cantidad_total_requisitos = configuracion.requisitosbecas.count() if configuracion.requisitosbecas else 0
                        if action_extra == 'requisitos_completos':
                            nombre_archivo = 'alumnos_migrantes_OCAS'
                            titulo_archivo = 'LISTADO DE ESTUDIANTES SELECCIONADOS POR SER ECUATORIANO MIGRANTE RETORNADO O DEPORTADO'
                        elif action_extra == 'requisitos_incompletos':
                            nombre_archivo = 'alumnos_migrantes_rechazados'
                            titulo_archivo = 'LISTADO DE ESTUDIANTES NO SELECCIONADOS POR SER ECUATORIANO MIGRANTE RETORNADO O DEPORTADO'

                font_style = XFStyle()
                font_style.font.bold = True
                font_style2 = XFStyle()
                font_style2.font.bold = False
                wb = Workbook(encoding='utf-8')
                # periodoactual = Periodo.objects.get(id=int(request.POST['periodoactual']))
                # periodovalida = Periodo.objects.get(id=int(request.POST['periodovalida']))
                periodoactual = periodosesion
                periodovalida = anterior
                # modalidad = int(request.POST['modalidad'])
                ws = wb.add_sheet('alumnos_migrantes')
                response = HttpResponse(content_type="application/ms-excel")
                response['Content-Disposition'] = f'attachment; filename={nombre_archivo}' + random.randint(1,
                                                                                                            10000).__str__() + '.xls'
                columns = [
                    (u"N.", 1500),
                    (u"ID_PREINSCRIPCION.", 1500),
                    (u"ORDEN", 1500),
                    (u"APELLIDOS Y NOMBRES", 12000),
                    (u"CARRERA", 8000),
                    (u"MODALIDAD", 8000),
                    (u"CEDULA", 3000),
                    (u"FECHA NACIMIENTO", 3000),
                    (u"PAIS RETORNO", 3000),
                    (u"AÑOS/MESES", 3000),
                    (u"FECHA DE RETORNO", 3000),
                    (u"PROMEDIO", 3000),
                    (u"ASISTENCIA", 3000),
                    (u"SESION", 3000),
                    (u"NIVEL", 3000),
                    (u"PARALELO", 3000),
                    (u"DIRECCION", 3000),
                    (u"TELEFONO", 4000),
                    (u"TELEFONO CONVENCIONAL", 4000),
                    (u"EMAIL", 5000),
                    (u"EMAIL INSTITUCIONAL", 5000),
                    (u"SEXO", 5000),
                    (u"PAIS", 5000),
                    (u"PROVINCIA", 5000),
                    (u"CANTON", 5000),
                    (u"DIRECCION 1", 5000),
                    (u"DIRECCION 2", 5000),
                    (u"SECTOR", 5000),
                    (u"GRUPO SOCIOECONOMICO", 5000),
                    (u"PERIODO ACTUAL", 8000),
                    (u"PERIDOO VALIDA", 8000),
                    (u"TIPO/ESTADO", 2500),
                ]

                requisitos = []
                if configuracion:
                    #Requisitos de tipos de becas
                    requisitos = configuracion.requisitosbecas.filter(visible=True, status=True)
                    for detallerequisitobeca in requisitos:
                        columns.append((f'({detallerequisitobeca.pk}) {detallerequisitobeca.requisitobeca.__str__()}', 10000))
                        if not action_extra:
                            columns.append((f'({detallerequisitobeca.pk}) OBSERVACIÓN', 5000))

                style_title = xlwt.easyxf(
                    'font: height 350, name Arial, colour_index black, bold on, italic on; align: wrap on, vert centre, horiz center;')
                style_title_2 = xlwt.easyxf(
                    'font: height 350, name Arial, colour_index black, bold on, italic on; align: wrap on, vert centre, horiz center;')
                ws.write_merge(0, 0, 0, len(columns), 'UNIVERSIDAD ESTATAL DE MILAGRO', style_title)
                ws.write_merge(1, 1, 0, len(columns), titulo_archivo, style_title_2)
                row_num = 3
                for col_num in range(len(columns)):
                    ws.write(row_num, col_num, columns[col_num][0], font_style)
                    ws.col(col_num).width = columns[col_num][1]
                date_format = xlwt.XFStyle()
                date_format.num_format_str = 'yyyy/mm/dd'
                if periodosesion.versionbeca is None or periodosesion.versionbeca == 1:
                    discapacitados = Matricula.objects.values_list('inscripcion__id', flat=True).filter(status=True,
                                                                                                        nivel__periodo=periodoactual,
                                                                                                        estado_matricula__in=[
                                                                                                            2, 3],
                                                                                                        retiradomatricula=False,
                                                                                                        inscripcion__persona__perfilinscripcion__tienediscapacidad=True,
                                                                                                        inscripcion__persona__perfilinscripcion__verificadiscapacidad=True,
                                                                                                        matriculagruposocioeconomico__tipomatricula=1).exclude(
                        inscripcion__carrera__coordinacion__id__in=[7, 9]).distinct().order_by("inscripcion__persona")
                    matriculados = Matricula.objects.filter(status=True, nivel__periodo=periodovalida,
                                                            estado_matricula__in=[2, 3],
                                                            retiradomatricula=False,
                                                            matriculagruposocioeconomico__tipomatricula=1,
                                                            inscripcion__id__in=discapacitados).exclude(
                        inscripcion__carrera__coordinacion__id__in=[7, 9]).distinct().order_by("inscripcion__persona")
                elif periodosesion.versionbeca == 2:
                    inscripciones = PreInscripcionBeca.objects.filter(periodo=periodosesion,
                                                                      becatipo_id=22).values_list('inscripcion_id',
                                                                                                  flat=True)
                    if action_extra == 'requisitos_completos':
                        inscripciones = inscripciones.filter(preinscripcionbecarequisito__cumplerequisito=True)\
                            .annotate(total_requisitos=Count('preinscripcionbecarequisito', filter=Q(status=True), distinct=True))\
                            .filter(total_requisitos=cantidad_total_requisitos)
                    elif action_extra == 'requisitos_incompletos':
                        pendientes = inscripciones.values_list('id', flat=True).filter(preinscripcionbecarequisito__cumplerequisito__isnull=True)
                        inscripciones = inscripciones.filter(preinscripcionbecarequisito__cumplerequisito=False).exclude(id__in=pendientes).distinct()

                    matriculados = Matricula.objects.filter(inscripcion__persona__migrantepersona__isnull=False,
                                                            inscripcion__persona__pais_id=1,
                                                            inscripcion__id__in=inscripciones.values_list('inscripcion_id',flat=True),
                                                            nivel__periodo__id=periodo.id).distinct().order_by(
                        "inscripcion__persona")
                row_num = 4
                i = 0
                for x in matriculados:
                    preinscripcion = PreInscripcionBeca.objects.filter(inscripcion=x.inscripcion, periodo=periodo, becatipo_id=22).first()
                    asignaturas = MateriaAsignada.objects.filter(status=True,
                                                                 retiramateria=False,
                                                                 matricula__inscripcion__id=x.inscripcion.id,
                                                                 matricula__nivel__periodo__id=anterior.id,
                                                                 materiaasignadaretiro__isnull=True).exclude(materia__asignaturamalla__malla_id__in=[353, 22])
                    verifica = 0
                    suma = 0
                    promedio = 0
                    sumasis = 0
                    total = asignaturas.count()
                    for m in asignaturas:
                        suma += m.notafinal
                        sumasis += m.asistenciafinal
                        if m.estado.id != 1:
                            verifica = 1
                            break
                    if suma > 0 and verifica == 0:
                        promedio = round(suma / total, 2)
                        asistencia = round(sumasis / total, 2)
                    #if verifica == 0:
                    campo1 = preinscripcion.id
                    campo2 = preinscripcion.orden
                    campo3 = x.inscripcion.persona.nombre_completo_inverso()
                    if x.inscripcion.carrera.mencion:
                        campo4 = x.inscripcion.carrera.nombre + ' CON MENCION EN  ' + x.inscripcion.carrera.mencion
                    else:
                        campo4 = x.inscripcion.carrera.nombre
                    campo5 = x.inscripcion.modalidad.__str__()
                    campo6 = x.inscripcion.persona.cedula
                    campo7 = x.inscripcion.persona.nacimiento
                    campo8 = x.inscripcion.persona.registro_migrante().paisresidencia.nombre
                    campo9 = u"%s - %s" % (x.inscripcion.persona.registro_migrante().anioresidencia,
                                           x.inscripcion.persona.registro_migrante().mesresidencia)
                    campo10 = x.inscripcion.persona.registro_migrante().fecharetorno
                    campo11 = promedio
                    campo12 = asistencia
                    campo13 = x.inscripcion.sesion.nombre
                    if x.nivelmalla:
                        campo14 = x.nivelmalla.nombre
                    else:
                        campo14 = 'Ninguno'
                    if x.paralelo:
                        campo15 = x.paralelo.nombre
                    else:
                        campo15 = 'Ninguno'
                    campo16 = x.inscripcion.persona.direccion_completa()
                    campo17 = x.inscripcion.persona.telefono
                    campo18 = x.inscripcion.persona.telefono_conv
                    campo19 = x.inscripcion.persona.email
                    campo20 = x.inscripcion.persona.emailinst
                    campo21 = x.inscripcion.persona.sexo.nombre
                    campo22 = x.inscripcion.persona.pais.nombre if x.inscripcion.persona.pais else ""
                    campo23 = x.inscripcion.persona.provincia.nombre if x.inscripcion.persona.provincia else ""
                    campo24 = x.inscripcion.persona.canton.nombre if x.inscripcion.persona.canton else ""
                    campo25 = x.inscripcion.persona.direccion if x.inscripcion.persona.direccion else ""
                    campo26 = x.inscripcion.persona.direccion2 if x.inscripcion.persona.direccion2 else ""
                    campo27 = x.inscripcion.persona.sector if x.inscripcion.persona.sector else ""
                    campo28 = str(
                        x.matriculagruposocioeconomico().nombre) if x.matriculagruposocioeconomico() else ""
                    campo29 = x.estado_renovacion_beca(becatipo, anterior)
                    i += 1
                    ws.write(row_num, 0, i, font_style2)
                    ws.write(row_num, 1, campo1, font_style2)
                    ws.write(row_num, 2, campo2, font_style2)
                    ws.write(row_num, 3, campo3, font_style2)
                    ws.write(row_num, 4, campo4, font_style2)
                    ws.write(row_num, 5, campo5, font_style2)
                    ws.write(row_num, 6, campo6, font_style2)
                    ws.write(row_num, 7, campo7, date_format)
                    ws.write(row_num, 8, campo8, font_style2)
                    ws.write(row_num, 9, campo9, font_style2)
                    ws.write(row_num, 10, campo10, date_format)
                    ws.write(row_num, 11, campo11, font_style2)
                    ws.write(row_num, 12, campo12, font_style2)
                    ws.write(row_num, 13, campo13, font_style2)
                    ws.write(row_num, 14, campo14, font_style2)
                    ws.write(row_num, 15, campo15, font_style2)
                    ws.write(row_num, 16, campo16, font_style2)
                    ws.write(row_num, 17, campo17, font_style2)
                    ws.write(row_num, 18, campo18, font_style2)
                    ws.write(row_num, 19, campo19, font_style2)
                    ws.write(row_num, 20, campo20, font_style2)
                    ws.write(row_num, 21, campo21, font_style2)
                    ws.write(row_num, 22, campo22, font_style2)
                    ws.write(row_num, 23, campo23, font_style2)
                    ws.write(row_num, 24, campo24, font_style2)
                    ws.write(row_num, 25, campo25, font_style2)
                    ws.write(row_num, 26, campo26, font_style2)
                    ws.write(row_num, 27, campo27, font_style2)
                    ws.write(row_num, 28, campo28, font_style2)
                    ws.write(row_num, 29, str(periodovalida), font_style2)
                    ws.write(row_num, 30, str(periodoactual), font_style2)
                    ws.write(row_num, 31, campo29, font_style2)

                    if requisitos:
                        col_num = len(columns) - requisitos.count()*(2 if not action_extra else 1)
                        for detallerequisitobeca in requisitos:
                            preinscriocionrequisitos = PreInscripcionBecaRequisito.objects.filter(preinscripcion=preinscripcion, detallerequisitobeca=detallerequisitobeca, status=True)
                            for preinsrequisito in preinscriocionrequisitos:
                                cumple = ''
                                cumplerequisito = preinsrequisito.cumplerequisito
                                if cumplerequisito:
                                   cumple = 'SI'
                                elif cumplerequisito == False:
                                   cumple = 'NO'
                                ws.write(row_num, col_num, cumple, font_style2)
                                if not action_extra:
                                    col_num += 1
                                    ws.write(row_num, col_num, '', font_style2) #Observación por columna
                            col_num += 1
                    row_num += 1
                wb.save(response)
                return response
            except Exception as ex:
                err = 'Error on line {}'.format(sys.exc_info()[-1].tb_lineno)
                msg = ex.__str__()
                msg = f'{msg} {err}'

                return JsonResponse({"result": "bad", "mensaje": u"Error: %s"%msg})

        elif action == 'deletesolicitud':
            try:
                if BecaSolicitud.objects.get(pk=int(request.POST['id']), status=True):
                    solicitud = BecaSolicitud.objects.get(pk=int(request.POST['id']), status=True)
                    solicitud.delete()
                    log(u'Elimino solicitud de beca: %s' % solicitud, request, "del")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'editbeca':
            try:
                becaasignacion = BecaAsignacion.objects.get(pk=int(request.POST['id']), status=True)
                f = BecaAsignacionForm(request.POST)
                if f.is_valid():
                    montomensual = null_to_decimal(f.cleaned_data['montomensual'], 2)
                    montobeneficio = null_to_decimal(f.cleaned_data['montobeneficio'], 2)
                    if montomensual <= 0 or montobeneficio <= 0:
                        return JsonResponse({"result": "bad", "mensaje": u"Ingrese montos mayores a cero."})
                    cantidadmeses = int(f.cleaned_data['cantidadmeses'])
                    total = null_to_decimal(montomensual * cantidadmeses, 2)
                    if montobeneficio != total:
                        return JsonResponse({"result": "bad", "mensaje": u"montos no cuadran."})
                    becaasignacion.montobeneficio = f.cleaned_data['montobeneficio']
                    becaasignacion.cantidadmeses = f.cleaned_data['cantidadmeses']
                    becaasignacion.montomensual = f.cleaned_data['montomensual']
                    becaasignacion.grupopago = f.cleaned_data['grupopago']
                    becaasignacion.tipo = f.cleaned_data['tipo']
                    becaasignacion.estadobeca = f.cleaned_data['estadobeca']
                    becaasignacion.save(request)
                    log(u'Edita Beca: %s' % becaasignacion, request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'deletebecaasignacion':
            try:
                if BecaAsignacion.objects.get(pk=int(request.POST['id']), status=True):
                    becaasignacion = BecaAsignacion.objects.get(pk=int(request.POST['id']), status=True)
                    becaasignacion.delete()
                    log(u'Elimino beca: %s' % becaasignacion, request, "del")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'importarbecados':
            try:
                f = ImportarBecaForm(request.POST, request.FILES)
                if f.is_valid():
                    nfile = request.FILES['archivo']
                    nfile._name = generar_nombre("importacion_becas_", nfile._name)
                    archivo = Archivo(nombre='IMPORTACION BECAS',
                                      fecha=datetime.now().date(),
                                      archivo=nfile,
                                      tipo_id=ARCHIVO_TIPO_GENERAL)
                    archivo.save(request)
                    workbook = xlrd.open_workbook(archivo.archivo.file.name)
                    sheet = workbook.sheet_by_index(0)
                    linea = 0
                    periodoescojido = f.cleaned_data['periodo']
                    becatipo = f.cleaned_data['becatipo']
                    observacion = f.cleaned_data['observacion']
                    tipo = f.cleaned_data['tipo']
                    grupopago = f.cleaned_data['grupopago']
                    for rowx in range(sheet.nrows):
                        puntosalva = transaction.savepoint()
                        try:
                            if linea > 0:
                                cols = sheet.row_values(rowx)
                                identificacion = str(cols[0]).strip().upper()
                                identificacion = identificacion.replace(".0", "")
                                if Persona.objects.values('id').filter(
                                        Q(cedula__icontains=identificacion) | Q(pasaporte__icontains=identificacion),
                                        status=True).exists():
                                    persona = Persona.objects.filter(
                                        Q(cedula__icontains=identificacion) | Q(pasaporte__icontains=identificacion),
                                        status=True)[0]
                                    if Inscripcion.objects.values('id').filter(status=True, persona=persona).exists():
                                        inscripciones = Inscripcion.objects.filter(status=True, persona=persona,
                                                                                   matricula__nivel__periodo=periodoescojido).order_by(
                                            'id')
                                        if inscripciones:
                                            inscripcion = inscripciones[0]
                                            if not BecaSolicitud.objects.values('id').filter(inscripcion=inscripcion,
                                                                                             becatipo=becatipo,
                                                                                             periodo=periodoescojido,
                                                                                             estado=2).exists():
                                                becasolicitud = BecaSolicitud(inscripcion=inscripcion,
                                                                              becatipo=becatipo,
                                                                              periodo=periodoescojido, estado=2,
                                                                              observacion=observacion)
                                                becasolicitud.save()
                                            else:
                                                becasolicitud = BecaSolicitud.objects.filter(inscripcion=inscripcion,
                                                                                             becatipo=becatipo,
                                                                                             periodo=periodoescojido,
                                                                                             estado=2)[0]

                                            for x in BecaRequisitos.objects.filter(status=True):
                                                becasolicituddetalle = BecaDetalleSolicitud(solicitud=becasolicitud,
                                                                                            requisito=x,
                                                                                            cumple=True,
                                                                                            estado=2,
                                                                                            observacion=u"APROBADO AUTOMÁTICO IMPORTACIÓN")
                                                becasolicituddetalle.save()
                                            if not BecaAsignacion.objects.values('id').filter(status=True,
                                                                                              solicitud=becasolicitud).exists():
                                                montomensual = null_to_decimal(cols[1], 2)
                                                cantidadmeses = int(cols[2])
                                                montobeneficio = null_to_decimal(montomensual * cantidadmeses, 2)
                                                asignacion = BecaAsignacion(solicitud=becasolicitud,
                                                                            montomensual=null_to_decimal(cols[1], 2),
                                                                            cantidadmeses=cantidadmeses,
                                                                            montobeneficio=montobeneficio,
                                                                            fecha=datetime.now().date(),
                                                                            activo=True, tipo=tipo, grupopago=grupopago)
                                                asignacion.save()
                            else:
                                linea += 1
                        except Exception as ex:
                            transaction.savepoint_rollback(puntosalva)
                            return JsonResponse({"result": "bad", "mensaje": u"Error al ingresar la linea: %s" % linea})
                    log(u'importo becas: %s %s %s %s' % (archivo, periodoescojido, becatipo, observacion), request,
                        "add")
                    return JsonResponse({'result': 'ok'})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', 'mensaje': u'Error al guardar los datos'})

        elif action == 'addbecado':
            try:
                f = BecaAsignacionManualForm(request.POST, request.FILES)
                if f.is_valid():
                    periodoescojido = periodosesion
                    becatipo = f.cleaned_data['becatipo']
                    observacion = f.cleaned_data['observacion']
                    inscripcion = Inscripcion.objects.get(id=f.cleaned_data['inscripcion'])
                    if not BecaSolicitud.objects.values('id').filter(inscripcion=inscripcion, becatipo=becatipo,
                                                                     periodo=periodoescojido, estado=2).exists():
                        becasolicitud = BecaSolicitud(inscripcion=inscripcion, becatipo=becatipo,
                                                      periodo=periodoescojido, estado=2,
                                                      observacion=observacion)
                        becasolicitud.save()
                    else:
                        becasolicitud = BecaSolicitud.objects.filter(inscripcion=inscripcion,
                                                                     becatipo=becatipo,
                                                                     periodo=periodoescojido, estado=2)[0]

                    for x in BecaRequisitos.objects.filter(status=True):
                        becasolicituddetalle = BecaDetalleSolicitud(solicitud=becasolicitud,
                                                                    requisito=x,
                                                                    cumple=True,
                                                                    estado=2,
                                                                    observacion=u"APROBADO AUTOMÁTICO AGREGACIÓN MANUAL DE %s" % persona.usuario)
                        becasolicituddetalle.save()
                    if not BecaAsignacion.objects.values('id').filter(status=True, solicitud=becasolicitud).exists():
                        asignacion = BecaAsignacion(solicitud=becasolicitud,
                                                    montomensual=f.cleaned_data['montomensual'],
                                                    cantidadmeses=f.cleaned_data['cantidadmeses'],
                                                    montobeneficio=f.cleaned_data['montobeneficio'],
                                                    tipo=f.cleaned_data['tipo'],
                                                    grupopago=f.cleaned_data['grupopago'],
                                                    fecha=periodoescojido.inicio,
                                                    activo=True)
                        asignacion.save()
                    log(u'adiciono becado: %s %s %s %s' % (inscripcion, periodoescojido, becatipo, observacion),
                        request, "add")
                    return JsonResponse({'result': 'ok'})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', 'mensaje': u'Error al guardar los datos'})

        elif action == 'addBecaTipoConfiguracion':
            try:
                idpb = None
                if 'idpb' in request.POST and request.POST['idpb']:
                    idpb = request.POST['idpb']
                    if not BecaPeriodo.objects.filter(pk=idpb, status=True).exists():
                        raise NameError(u"No se encontro el periodo de beca")
                else:
                    raise NameError(u"No se encontro el periodo de beca")
                becaperiodo = BecaPeriodo.objects.get(pk=idpb)
                f = BecaTipoConfiguracionForm(request.POST)
                if not f.is_valid():
                    raise NameError(u"Formlario incorrecto")
                becatipo = f.cleaned_data['becatipo']
                becamonto = f.cleaned_data['becamonto']
                becameses = f.cleaned_data['becameses']
                if BecaTipoConfiguracion.objects.filter(becaperiodo=becaperiodo, becatipo=becatipo).exists():
                    raise NameError(u"Tipo de beca ya configurado en el periodo")
                x = BecaTipoConfiguracion(becaperiodo=becaperiodo,
                                          becatipo=becatipo,
                                          becamonto=becamonto,
                                          becameses=becameses)

                x.save(request)
                for re in f.cleaned_data['requisitosbecas']:
                    x.requisitosbecas.add(re)
                    x.save()
                for re in f.cleaned_data['documentos']:
                    x.documentos.add(re)
                    x.save()
                log(u'Adiciono configuración de tipo de beca %s del periodo %s' % (
                becatipo.nombre, becaperiodo.periodo.nombre), request, "add")
                return JsonResponse({'result': 'ok'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', 'mensaje': u'Error al guardar los datos. %s' % ex})

        elif action == 'editBecaTipoConfiguracion':
            try:
                idpb = None
                if 'idpb' in request.POST and request.POST['idpb']:
                    idpb = request.POST['idpb']
                    if not BecaPeriodo.objects.filter(pk=idpb, status=True).exists():
                        raise NameError(u"No se encontro el periodo de beca")
                else:
                    raise NameError(u"No se encontro el periodo de beca")
                x = BecaTipoConfiguracion.objects.get(pk=int(request.POST['id']))
                becaperiodo = BecaPeriodo.objects.get(pk=idpb)
                f = BecaTipoConfiguracionForm(request.POST)
                if not f.is_valid():
                    raise NameError(u"Formlario incorrecto")
                x.becatipo = f.cleaned_data['becatipo']
                x.becamonto = f.cleaned_data['becamonto']
                x.becameses = f.cleaned_data['becameses']
                x.requisitosbecas.clear()
                for data in f.cleaned_data['requisitosbecas']:
                    x.requisitosbecas.add(data)
                for data in f.cleaned_data['documentos']:
                    x.documentos.add(data)
                x.save(request)
                log(u'Modifico configuración de tipo de beca %s del periodo %s' % (
                x.becatipo.nombre, x.becaperiodo.periodo.nombre), request, "edit")
                return JsonResponse({'result': 'ok'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', 'mensaje': u'Error al guardar los datos'})

        elif action == 'reportebecas':
            mensaje = "Problemas al generar reporte."
            try:
                becas = BecaAsignacion.objects.filter(status=True)
                if int(request.POST['periodo']) != 0:
                    data['periodo'] = periodoselec = Periodo.objects.get(id=int(request.POST['periodo']))
                    becas = becas.filter(solicitud__periodo=periodoselec)
                if int(request.POST['carrera']) != 0:
                    data['carrera'] = carrera = Carrera.objects.get(id=int(request.POST['carrera']))
                    becas = becas.filter(solicitud__inscripcion__carrera=carrera)
                if request.POST['fechainicio'] and request.POST['fechafin']:
                    data['fechainicio'] = inicio = convertir_fecha(request.POST['fechainicio'])
                    data['fechafin'] = fin = convertir_fecha(request.POST['fechafin'])
                    becas = becas.filter(fecha__gte=inicio, fecha__lte=fin)
                if int(request.POST['tipobeca9']) != 0:
                    data['tipobeca9'] = tipobeca9 = BecaTipo.objects.get(id=int(request.POST['tipobeca9']))
                    becas = becas.filter(solicitud__becatipo=tipobeca9)
                if int(request.POST['tipobeca91']) != 0:
                    data['tipobeca91'] = tipobeca91 = int(request.POST['tipobeca91'])
                    becas = becas.filter(tipo=tipobeca91)
                data['becas'] = becas.order_by('solicitud__inscripcion__persona__apellido1',
                                               'solicitud__inscripcion__persona__apellido2',
                                               'solicitud__inscripcion__persona__nombres')
                data['valormensual'] = null_to_numeric(becas.aggregate(valor=Sum('montomensual'))['valor'])
                data['total'] = null_to_numeric(becas.aggregate(valor=Sum('montobeneficio'))['valor'])
                distributivos = DistributivoPersona.objects.filter(denominacionpuesto=547, status=True,
                                                                   estadopuesto_id=1)
                data['distributivo'] = distributivos[0] if distributivos.values('id').exists() else None
                cantidadmeses = becas.aggregate(cantidadmeses=Max("cantidadmeses"))['cantidadmeses']
                data['totalcantidadmeses'] = cantidadmeses
                contador_letra = ['PRIMER MES', 'SEGUNDO MES', 'TERCERO MES', 'CUARTO MES', 'QUINTO MES', 'SEXTO MES',
                                  'SEPTIMO MES', 'OCTAVO MES', 'NOVENO MES']
                data['lista_meses'] = contador_letra[:cantidadmeses]
                return conviert_html_to_pdf('adm_becas/reportebecas.html', {'pagesize': 'A4', 'data': data})
            except Exception as ex:
                return HttpResponseRedirect("/adm_becas?info=%s" % mensaje)

        elif action == 'notificarsolicitud':
            try:
                asunto = u"NOTIFICACION DE SOLICITUD BECA"
                id = int(request.POST['id'])
                becado = BecaSolicitud.objects.get(id=id, periodo=periodo, status=True)
                send_html_mail(asunto, "emails/notificarbecado.html",
                               {'sistema': request.session['nombresistema'],
                                'alumno': becado.inscripcion.persona.nombre_completo_inverso()},
                               becado.inscripcion.persona.lista_emails_envio(), [],
                               cuenta=CUENTAS_CORREOS[0][1])
                return JsonResponse({'result': 'ok'})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al notoficar."})

        elif action == 'notificarporsistema':
            try:
                becado = BecaAsignacion.objects.get(id=int(request.POST['id']), solicitud__periodo=periodo, status=True)
                becado.notificar = True
                becado.save(request)
                return JsonResponse({'result': 'ok'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al notoficar."})

        elif action == 'matriz':
            try:
                font_style = XFStyle()
                font_style.font.bold = True
                font_style2 = XFStyle()
                font_style2.font.bold = False
                wb = Workbook(encoding='utf-8')
                ws = wb.add_sheet('becados')
                response = HttpResponse(content_type="application/ms-excel")
                response['Content-Disposition'] = 'attachment; filename=becados' + random.randint(1,
                                                                                                  10000).__str__() + '.xls'
                columns = [
                    (u"CODIGO_IES", 2000),
                    (u"CODIGO_CARRERA", 2000),
                    (u"TIPO_IDENTIFICACION", 2000),
                    (u"IDENTIFICACION", 2000),
                    (u"CODIGO_BECA", 2000),
                    (u"ANIO", 2000),
                    (u"FECHA_INICIO_PERIODO", 2000),
                    (u"FECHA_FIN_PERIODO", 2000),
                    (u"TIPO_AYUDA", 2000),
                    (u"MOTIVO_BECA", 12000),
                    (u"OTRO_MOTIVO", 2000),
                    (u"MONTO_RECIBIDO", 2000),
                    (u"PORCENTAJE_VALOR_ARANCEL", 2000),
                    (u"PORCENTAJE_MANUTENCION", 2000),
                    (u"TIPO_FINANCIAMIENTO", 12000),
                    (u"ALUMNO", 12000)
                ]
                row_num = 0
                for col_num in range(len(columns)):
                    ws.write(row_num, col_num, columns[col_num][0], font_style)
                    ws.col(col_num).width = columns[col_num][1]
                date_format = xlwt.XFStyle()
                date_format.num_format_str = 'yyyy/mm/dd'
                row_num = 1
                resultados = BecaAsignacion.objects.filter(status=True, solicitud__periodo=periodo,
                                                           solicitud__status=True)
                for x in resultados:
                    campo1 = '1024'
                    if (x.solicitud.inscripcion.mi_malla()):
                        campo2 = x.solicitud.inscripcion.mi_malla().codigo
                    else:
                        campo2 = ''
                    campo3 = str(x.solicitud.inscripcion.persona.tipo_identificacion_completo())
                    campo4 = x.solicitud.inscripcion.persona.identificacion()
                    if x.solicitud.inscripcion.matricula_periodo(periodo):
                        campo5 = str(x.solicitud.inscripcion.matricula_periodo(periodo).id) + "-" + str(
                            x.solicitud.becatipo_id)
                    else:
                        campo5 = ''
                    campo6 = periodo.anio
                    campo7 = str(periodo.inicio)
                    campo8 = str(periodo.fin)
                    campo9 = 'BECA COMPLETA'
                    campo10 = x.solicitud.becatipo.nombre if x.solicitud.becatipo else ''
                    campo11 = ''
                    campo12 = str(x.montobeneficio)
                    campo13 = ''
                    campo14 = ''
                    campo15 = 'RECURSOS FISCALES'
                    ws.write(row_num, 0, campo1, font_style2)
                    ws.write(row_num, 1, campo2, font_style2)
                    ws.write(row_num, 2, campo3, font_style2)
                    ws.write(row_num, 3, campo4, font_style2)
                    ws.write(row_num, 4, campo5, font_style2)
                    ws.write(row_num, 5, campo6, font_style2)
                    ws.write(row_num, 6, campo7, font_style2)
                    ws.write(row_num, 7, campo8, font_style2)
                    ws.write(row_num, 8, campo9, font_style2)
                    ws.write(row_num, 9, campo10, font_style2)
                    ws.write(row_num, 10, campo11, font_style2)
                    ws.write(row_num, 11, campo12, font_style2)
                    ws.write(row_num, 12, campo13, font_style2)
                    ws.write(row_num, 13, campo14, font_style2)
                    ws.write(row_num, 14, campo15, font_style2)
                    ws.write(row_num, 15, '%s' % x.solicitud.inscripcion.persona.nombre_completo(), font_style2)
                    row_num += 1
                wb.save(response)
                return response
            except Exception as ex:
                pass

        elif action == 'listadobecados':
            try:
                font_style = XFStyle()
                font_style.font.bold = True
                font_style2 = XFStyle()
                font_style2.font.bold = False
                wb = Workbook(encoding='utf-8')
                ws = wb.add_sheet('becados')
                response = HttpResponse(content_type="application/ms-excel")
                response['Content-Disposition'] = 'attachment; filename=becados' + random.randint(1,
                                                                                                  10000).__str__() + '.xls'
                columns = [
                    (u"CODIGO_IES", 2000),
                    (u"CODIGO_CARRERA", 2000),
                    (u"TIPO_IDENTIFICACION", 2000),
                    (u"IDENTIFICACION", 2000),
                    (u"CODIGO_BECA", 2000),
                    (u"ANIO", 2000),
                    (u"FECHA_INICIO_PERIODO", 2000),
                    (u"FECHA_FIN_PERIODO", 2000),
                    (u"TIPO_AYUDA", 2000),
                    (u"MOTIVO_BECA", 12000),
                    (u"OTRO_MOTIVO", 2000),
                    (u"MONTO_RECIBIDO", 2000),
                    (u"PORCENTAJE_VALOR_ARANCEL", 2000),
                    (u"PORCENTAJE_MANUTENCION", 2000),
                    (u"TIPO_FINANCIAMIENTO", 12000),
                    (u"NOMBRE", 12000),
                    (u"CARRERA", 12000),
                    (u"DIRECCION", 12000),
                    (u"GRUPO SOCIOECONOMICO", 12000),
                    (u"SEXO", 12000),
                    (u"PROMEDIO", 12000)

                ]
                row_num = 0
                for col_num in range(len(columns)):
                    ws.write(row_num, col_num, columns[col_num][0], font_style)
                    ws.col(col_num).width = columns[col_num][1]
                date_format = xlwt.XFStyle()
                date_format.num_format_str = 'yyyy/mm/dd'
                row_num = 1
                resultados = BecaAsignacion.objects.filter(status=True, solicitud__periodo=periodo,
                                                           solicitud__status=True)
                for x in resultados:
                    campo1 = '1024'
                    if (x.solicitud.inscripcion.mi_malla()):
                        campo2 = x.solicitud.inscripcion.mi_malla().codigo
                    else:
                        campo2 = ''
                    campo3 = str(x.solicitud.inscripcion.persona.tipo_identificacion_completo())
                    campo4 = x.solicitud.inscripcion.persona.identificacion()
                    if x.solicitud.inscripcion.matricula_periodo(periodo):
                        campo5 = str(x.solicitud.inscripcion.matricula_periodo(periodo).id) + "-" + str(
                            x.solicitud.becatipo_id)
                    else:
                        campo5 = ''
                    campo6 = periodo.anio
                    campo7 = str(periodo.inicio)
                    campo8 = str(periodo.fin)
                    campo9 = 'BECA COMPLETA'
                    campo10 = x.solicitud.becatipo.nombre
                    campo11 = ''
                    campo12 = str(x.montobeneficio)
                    campo13 = ''
                    campo14 = ''
                    campo15 = 'RECURSOS FISCALES'
                    campo16 = x.solicitud.inscripcion.persona.nombre_completo_inverso()
                    campo17 = x.solicitud.inscripcion.carerra
                    campo18 = x.solicitud.inscripcion.persona.direccion_completa()
                    campo19 = 'BAJO'
                    campo20 = x.solicitud.inscripcion.persona.sexo
                    campo21 = 'PROMEDIO'
                    ws.write(row_num, 0, campo1, font_style2)
                    ws.write(row_num, 1, campo2, font_style2)
                    ws.write(row_num, 2, campo3, font_style2)
                    ws.write(row_num, 3, campo4, font_style2)
                    ws.write(row_num, 4, campo5, font_style2)
                    ws.write(row_num, 5, campo6, font_style2)
                    ws.write(row_num, 6, campo7, font_style2)
                    ws.write(row_num, 7, campo8, font_style2)
                    ws.write(row_num, 8, campo9, font_style2)
                    ws.write(row_num, 9, campo10, font_style2)
                    ws.write(row_num, 10, campo11, font_style2)
                    ws.write(row_num, 11, campo12, font_style2)
                    ws.write(row_num, 12, campo13, font_style2)
                    ws.write(row_num, 13, campo14, font_style2)
                    ws.write(row_num, 14, campo15, font_style2)
                    ws.write(row_num, 15, campo16, font_style2)
                    ws.write(row_num, 16, campo17, font_style2)
                    ws.write(row_num, 17, campo18, font_style2)
                    ws.write(row_num, 18, campo19, font_style2)
                    ws.write(row_num, 19, campo20, font_style2)
                    row_num += 1
                wb.save(response)
                return response
            except Exception as ex:
                pass

        elif action == 'detallecalculobeca':
            try:
                id = int(request.POST['id']) if 'id' in request.POST and request.POST['id'] else None
                nota_ref = float(request.POST['promedio']) if 'promedio' in request.POST and request.POST[
                    'promedio'] else 0.0
                desviacion = float(request.POST['desviacion']) if 'desviacion' in request.POST and request.POST[
                    'desviacion'] else 0.0
                numbecados = int(request.POST['numbecados']) if 'numbecados' in request.POST and request.POST[
                    'numbecados'] else 0
                porbecados = float(request.POST['porbecados']) if 'porbecados' in request.POST and request.POST[
                    'porbecados'] else 0.0
                if not Malla.objects.filter(pk=id).exclude(pk__in=[22, 353]).exists() or not nota_ref or not desviacion:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})
                malla = Malla.objects.get(pk=id)
                nivelmalla_ids = AsignaturaMalla.objects.values_list('nivelmalla_id', flat=False).filter(status=True,
                                                                                                         malla=malla)
                asignaturamalla_ids = AsignaturaMalla.objects.values_list('id', flat=False).filter(status=True,
                                                                                                   malla=malla)
                niveles = NivelMalla.objects.filter(pk__in=nivelmalla_ids)
                # asignaturasmalla = AsignaturaMalla.objects.filter(status=True, malla=malla)
                inscripcion_ids = Matricula.objects.values_list('inscripcion_id', flat=False).filter(
                    nivel__periodo=periodo, inscripcion__carrera=malla.carrera, status=True, retiradomatricula=False)
                aData = []
                totalbecados = 0
                for nivel in niveles:
                    num_becados = 0
                    inscripciones = Inscripcion.objects.filter(pk__in=inscripcion_ids, inscripcionnivel__nivel=nivel)
                    for inscripcion in inscripciones:
                        promedio = null_to_decimal(RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                                  asignaturamalla_id__in=asignaturamalla_ids).aggregate(
                            promedio=Avg('nota'))['promedio'], 2)
                        if promedio >= nota_ref:
                            num_becados += 1
                    aData.append({"nivel_id": nivel.id,
                                  "nivel": nivel.nombre,
                                  "num_estudiantes": inscripciones.count(),
                                  "num_becados": num_becados})
                    totalbecados += num_becados
                data['totalbecados'] = totalbecados
                data['numbecados'] = numbecados
                data['porbecados'] = porbecados
                data['desviacion'] = desviacion
                data['nota_ref'] = nota_ref
                data['aData'] = aData
                template = get_template("adm_becas/detallecalculobeca.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'html': json_content})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'alumnoscalculobeca':
            try:
                id = int(request.POST['id']) if 'id' in request.POST and request.POST['id'] else None
                nota_ref = float(request.POST['promedio']) if 'promedio' in request.POST and request.POST[
                    'promedio'] else 0.0
                desviacion = float(request.POST['desviacion']) if 'desviacion' in request.POST and request.POST[
                    'desviacion'] else 0.0
                numbecados = int(request.POST['numbecados']) if 'numbecados' in request.POST and request.POST[
                    'numbecados'] else 0
                porbecados = float(request.POST['porbecados']) if 'porbecados' in request.POST and request.POST[
                    'porbecados'] else 0.0
                if not Malla.objects.filter(pk=id).exclude(pk__in=[22, 353]).exists() or not nota_ref or not desviacion:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})
                malla = Malla.objects.get(pk=id)
                carrera = Carrera.objects.filter(pk=malla.carrera.id)
                asignaturamalla_ids = AsignaturaMalla.objects.values_list('id', flat=False).filter(status=True,
                                                                                                   malla=malla)
                inscripcion_ids = Matricula.objects.values_list('inscripcion_id', flat=False).filter(
                    nivel__periodo=periodo, inscripcion__carrera=malla.carrera, status=True, retiradomatricula=False)
                aData = []
                inscripciones = Inscripcion.objects.filter(pk__in=inscripcion_ids)
                for inscripcion in inscripciones:
                    promedio = null_to_decimal(RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                              asignaturamalla_id__in=asignaturamalla_ids).aggregate(
                        promedio=Avg('nota'))['promedio'], 2)
                    if promedio >= nota_ref:
                        nivel = inscripcion.inscripcionnivel_set.filter(status=True).first()
                        periodo_id = TarjetaRegistroAcademico.objects.values_list('periodo_id', flat=True).filter(
                            inscripcion=inscripcion).order_by('periodo__inicio')
                        periodoinicio = None
                        if periodo_id.exists():
                            periodoinicio = Periodo.objects.get(pk=periodo_id.first())
                        aData.append({"inscripcion_id": inscripcion.id,
                                      "inscripcion": inscripcion.persona,
                                      "promedio": promedio,
                                      "nivel": nivel,
                                      "periodo": periodoinicio.nombre if periodoinicio else "S/R"})

                ord_aData = sorted(aData, key=lambda promedio: promedio['promedio'], reverse=True)
                data['numbecados'] = numbecados
                data['porbecados'] = porbecados
                data['desviacion'] = desviacion
                data['nota_ref'] = nota_ref
                data['aData'] = ord_aData
                template = get_template("adm_becas/alumnoscalculobeca.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'html': json_content})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'generarprecandidatos':
            try:
                GenerateBackground(request=request, data=data, periodoactual=periodo, periodoanterior=anterior).start()
                return JsonResponse({"result": "ok", "mensaje": u"Se ha procedido a ejecutar el proceso, en cuanto este se procedera a notificar"})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al generar los datos.%s " % ex})

        elif action == 'generarprecandidatosdesviacionestandarnew':
            try:
                GenerateBackground(request=request, data=data, periodoactual=periodo, periodoanterior=anterior).start()
                return JsonResponse({"result": "ok", "mensaje": u"Se ha procedido a ejecutar el proceso, en cuanto este se procedera a notificar"})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al generar los datos.%s " % ex})

        elif action == 'generarsolicitudesbecas':
            try:
                GenerateBackground(request=request, data=data, periodoactual=periodo, periodoanterior=anterior).start()
                return JsonResponse({"result": "ok", "mensaje": u"Se ha procedido a ejecutar el proceso, en cuanto este se procedera a notificar"})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al generar los datos.%s " % ex})

        elif action == 'generate_reporte_pendientes_pago_becas_financiero':
            try:
                noti = Notificacion(cuerpo='Generación de reporte de excel en progreso',
                                    titulo='Excel reporte de pendiente de pago financiero',
                                    destinatario=persona,
                                    url='',
                                    prioridad=1, app_label='SGA',
                                    fecha_hora_visible=datetime.now() + timedelta(days=1), tipo=2,
                                    en_proceso=True)
                noti.save(request)
                GenerateBackground(request=request, data=data, periodoactual=periodo, periodoanterior=anterior, noti=noti).start()
                return JsonResponse({"result": "ok",
                                     "mensaje": u"Se ha procedido a ejecutar el proceso, en cuanto este se procedera a notificar",
                                     "btn_notificaciones": traerNotificaciones(request, data, persona)})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al generar los datos.%s " % ex})

        elif action == 'restaurarsolicitud':
            try:
                id = int(request.POST['id']) if 'id' in request.POST and request.POST['id'] else None
                observacion = request.POST['observacion'] if 'observacion' in request.POST and request.POST[
                    'observacion'] else None
                if not BecaSolicitud.objects.filter(pk=id).exists() or not observacion:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})
                becasolicitud = BecaSolicitud.objects.get(pk=id)
                becasolicitud.becaaceptada = 1
                if BecaAsignacion.objects.filter(solicitud=becasolicitud):
                    becasolicitud.becaaceptada = 2
                    becaasignacion = BecaAsignacion.objects.get(solicitud=becasolicitud)
                    becaasignacion.status = True
                    becaasignacion.estadobeca = None
                    becaasignacion.personarevisacontrato = None
                    becaasignacion.estadorevisioncontrato = None
                    becaasignacion.personarevisadocumento = None
                    becaasignacion.save(request)
                becasolicitud.save(request)
                becarecorrido = BecaSolicitudRecorrido(solicitud=becasolicitud,
                                                       fecha=datetime.now().date(),
                                                       observacion=f"SOLICITUD AUTOMÁTICA",
                                                       estado=1)
                becarecorrido.save(request)
                becarecorrido = BecaSolicitudRecorrido(solicitud=becasolicitud,
                                                       fecha=datetime.now().date(),
                                                       observacion=f"EN REVISION",
                                                       estado=4)
                becarecorrido.save(request)
                becarecorrido = BecaSolicitudRecorrido(solicitud=becasolicitud,
                                                       fecha=datetime.now().date(),
                                                       observacion=f"PENDIENTE DE ACEPTACIÓN O RECHAZO",
                                                       estado=2)
                becarecorrido.save(request)
                becarecorrido = BecaSolicitudRecorrido(solicitud=becasolicitud,
                                                       fecha=datetime.now().date(),
                                                       observacion=f"BECA ACEPTADA. %s" % observacion.upper(),
                                                       estado=6)
                becarecorrido.save(request)
                messages.add_message(request, messages.SUCCESS, f'Se guardo exitosamente')
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'reporte_solicitudes_pago':
            try:
                becassolicitudpago = SolicitudPagoBecaDetalle.objects.filter(status=True,asignacion__solicitud__periodo=periodo).distinct().order_by(
                    'asignacion__solicitud__inscripcion__persona__apellido1',
                    'asignacion__solicitud__inscripcion__persona__apellido2',
                    'asignacion__solicitud__inscripcion__persona__nombres')

                if becassolicitudpago:
                    fuentenormal = easyxtitulo = easyxf(
                        'font: name Times New Roman, color-index black, bold on , height 350; alignment: horiz centre')
                    titulo = easyxf(
                        'font: name Times New Roman, color-index black, bold on , height 350; alignment: horiz centre')
                    titulo2 = easyxf(
                        'font: name Times New Roman, color-index black, bold on , height 250; alignment: horiz centre')
                    fuentecabecera = easyxf(
                        'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
                    fuentenormal = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
                    fuentenormalwrap = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
                    fuentenormalneg = easyxf(
                        'font: name Verdana, color-index black, bold on , height 150; borders: left thin, right thin, top thin, bottom thin')
                    fuentenormalnegrell = easyxf(
                        'font: name Verdana, color-index black, bold on , height 150; borders: left thin, right thin, top thin, bottom thin; pattern: pattern solid, fore_colour gray25')
                    fuentenormalwrap.alignment.wrap = True
                    fuentenormalcent = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz centre')
                    fuentemoneda = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right',
                        num_format_str=' "$" #,##0.00')
                    fuentemonedaneg = easyxf(
                        'font: name Verdana, color-index black, bold on, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right; pattern: pattern solid, fore_colour gray25',
                        num_format_str=' "$" #,##0.00')
                    fuentefecha = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz center',
                        num_format_str='yyyy-mm-dd')
                    fuentenumerodecimal = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right',
                        num_format_str='#,##0.00')
                    fuentenumeroentero = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right')

                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False

                    output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media', 'becaspregrado'))
                    try:
                        os.stat(output_folder)
                    except:
                        os.mkdir(output_folder)

                    libdestino = xlwt.Workbook()
                    hojadestino = libdestino.add_sheet("Reporte")

                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=solicitud_pago_beca_' + random.randint(1,
                                                                                                                   10000).__str__() + '.xls'
                    nombre = "SOLICITUD_PAGO_BECA_" + datetime.now().strftime('%Y%m%d_%H%M%S') + ".xls"
                    filename = os.path.join(output_folder, nombre)
                    ruta = "media/becaspregrado/" + nombre

                    hojadestino.write_merge(0, 0, 0, 14, 'UNIVERSIDAD ESTATAL DE MILAGRO', titulo)
                    hojadestino.write_merge(1, 1, 0, 14, 'DIRECCIÓN DE BIENESTAR UNIVERSITARIO', titulo2)
                    hojadestino.write_merge(2, 2, 0, 14, 'REPORTE SOLICITUDES DE PAGO PARA BECAS', titulo2)
                    hojadestino.write_merge(3, 3, 0, 14, 'PERIODO ' + periodo.nombre, titulo2)

                    fila = 5

                    columnas = [
                        (u"NÚMERO REPORTE", 2000),
                        (u"FECHA SOL.PAGO", 3000),
                        (u"TIPO BECA", 8000),
                        (u"MONTO BECA", 2500),
                        (u"NOMBRES COMPLETOS", 10000),
                        (u"IDENTIFICACIÓN", 4100),
                        (u"FACULTAD", 5000),
                        (u"CARRERA", 5000),
                        (u"MODALIDAD", 5000),
                        (u"NIVEL", 3000),
                        (u"CORREO PERSONAL", 6000),
                        (u"CORREO UNEMI", 6000),
                        (u"TELÉFONO", 3000),
                        (u"CELULAR", 3000),
                        (u"OPERADORA", 3000),
                        (u"NUEVA/RENV", 3000),
                        (u"NUM CUENTA", 4000),
                        (u"TIPO CUENTA", 3000),
                        (u"ID TIPO CUENTA", 4000),
                        (u"BANCO", 6000),
                        (u"CODIGO BANCO", 2000),
                        (u"ESTADO CUENTA", 4000)
                    ]

                    for col_num in range(len(columnas)):
                        hojadestino.write(fila, col_num, columnas[col_num][0], fuentecabecera)
                        hojadestino.col(col_num).width = columnas[col_num][1]

                    fila = 6
                    for solicitudpago in becassolicitudpago:
                        personasol = solicitudpago.asignacion.solicitud.inscripcion.persona
                        beca = solicitudpago.asignacion
                        matricula = None
                        #(MC) COD1368 - SE AGREGO PARA PODER AGREGAR VARIOS CAMPOS AL REPORTE
                        nasig = solicitudpago.asignacion
                        bancapersona = personasol.cuentabancariapersona_set.filter(activapago=True, verificado=True).first()
                        if not bancapersona:
                            bancapersona = None
                            tipobanco = None
                            banco = None
                        else:
                            tipobanco = bancapersona.tipocuentabanco
                            banco = bancapersona.banco
                            #FIN (MC) COD1368

                        if beca.solicitud.inscripcion.matricula_periodo_actual(periodo):
                            matricula = beca.solicitud.inscripcion.matricula_periodo_actual(periodo)[0]

                        hojadestino.write(fila, 0, str(solicitudpago.solicitudpago.numerosolicitud).zfill(5),fuentenormal)
                        hojadestino.write(fila, 1, solicitudpago.solicitudpago.fecha, fuentefecha)
                        hojadestino.write(fila, 2, beca.solicitud.becatipo.nombre.upper(), fuentenormal)
                        hojadestino.write(fila, 3, beca.montobeneficio, fuentemoneda)
                        hojadestino.write(fila, 4, personasol.nombre_completo_inverso(), fuentenormal)
                        hojadestino.write(fila, 5, personasol.identificacion(), fuentenormal)
                        hojadestino.write(fila, 6, str(matricula.nivel.coordinacion()) if matricula else '',fuentenormal)
                        hojadestino.write(fila, 7, str(matricula.inscripcion.carrera) if matricula else '',fuentenormal)
                        hojadestino.write(fila, 8, str(matricula.inscripcion.modalidad) if matricula else '',fuentenormal)
                        hojadestino.write(fila, 9, matricula.nivelmalla.nombre if matricula else '', fuentenormal)
                        hojadestino.write(fila, 10, personasol.email, fuentenormal)
                        hojadestino.write(fila, 11, personasol.emailinst, fuentenormal)
                        hojadestino.write(fila, 12, personasol.telefono_conv, fuentenormal)
                        hojadestino.write(fila, 13, personasol.telefono, fuentenormal)
                        hojadestino.write(fila, 14, personasol.get_tipocelular_display() if personasol.tipocelular else '',fuentenormal)
                        # (MC) -CAMPOS ADICIONALES A COD1368
                        hojadestino.write(fila, 15, nasig.get_tipo_display() if nasig.tipo else 'N/A', fuentenormal)
                        hojadestino.write(fila, 16, bancapersona.numero if bancapersona else '',fuentenormal)
                        hojadestino.write(fila, 17,tipobanco.nombre if bancapersona else '',fuentenormal)
                        hojadestino.write(fila, 18, tipobanco.id if bancapersona else '',fuentenormal)
                        hojadestino.write(fila, 19, banco.nombre if bancapersona else '',fuentenormal)
                        hojadestino.write(fila, 20, banco.codigo if bancapersona else '',fuentenormal)
                        hojadestino.write(fila, 21, bancapersona.get_estadorevision_display() if bancapersona else '',fuentenormal)
                        # (MC) -FIN COD1368

                        print(fila - 5)
                        fila += 1

                    libdestino.save(filename)
                    return JsonResponse({'result': 'ok', 'archivo': ruta})
                else:
                    return JsonResponse({'result': 'bad', 'mensaje': 'No existen registros para generar el reporte'})
            except Exception as ex:
                msg = ex.__str__()
                return JsonResponse({"result": "bad", "mensaje": u"Error al generar el reporte. [%s]" % msg})

        elif action == 'reporte_becas_pagadas':
            try:
                estado = int(request.POST['estado'])
                if estado == 1 or estado == 2:
                    estadopago = (estado == 1)
                    tituloreporte = "REPORTE DE BECAS PAGADAS" if estadopago else "REPORTE DE BECAS NO PAGADAS"
                    becassolicitudpago = SolicitudPagoBecaDetalle.objects.filter(status=True,
                                                                                 asignacion__solicitud__periodo=periodo,
                                                                                 pagado=estadopago).distinct().order_by(
                        'asignacion__solicitud__inscripcion__persona__apellido1',
                        'asignacion__solicitud__inscripcion__persona__apellido2',
                        'asignacion__solicitud__inscripcion__persona__nombres')
                else:
                    tituloreporte = "REPORTE DE BECAS PAGADAS Y NO PAGADAS"
                    becassolicitudpago = SolicitudPagoBecaDetalle.objects.filter(status=True,
                                                                                 asignacion__solicitud__periodo=periodo).distinct().order_by(
                        'asignacion__solicitud__inscripcion__persona__apellido1',
                        'asignacion__solicitud__inscripcion__persona__apellido2',
                        'asignacion__solicitud__inscripcion__persona__nombres')

                if becassolicitudpago:
                    fuentenormal = easyxtitulo = easyxf(
                        'font: name Times New Roman, color-index black, bold on , height 350; alignment: horiz centre')
                    titulo = easyxf(
                        'font: name Times New Roman, color-index black, bold on , height 350; alignment: horiz centre')
                    titulo2 = easyxf(
                        'font: name Times New Roman, color-index black, bold on , height 250; alignment: horiz centre')
                    fuentecabecera = easyxf(
                        'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
                    fuentenormal = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
                    fuentenormalwrap = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
                    fuentenormalneg = easyxf(
                        'font: name Verdana, color-index black, bold on , height 150; borders: left thin, right thin, top thin, bottom thin')
                    fuentenormalnegrell = easyxf(
                        'font: name Verdana, color-index black, bold on , height 150; borders: left thin, right thin, top thin, bottom thin; pattern: pattern solid, fore_colour gray25')
                    fuentenormalwrap.alignment.wrap = True
                    fuentenormalcent = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz centre')
                    fuentemoneda = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right',
                        num_format_str=' "$" #,##0.00')
                    fuentemonedaneg = easyxf(
                        'font: name Verdana, color-index black, bold on, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right; pattern: pattern solid, fore_colour gray25',
                        num_format_str=' "$" #,##0.00')
                    fuentefecha = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz center',
                        num_format_str='yyyy-mm-dd')
                    fuentenumerodecimal = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right',
                        num_format_str='#,##0.00')
                    fuentenumeroentero = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right')

                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False

                    output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media', 'becaspregrado'))
                    try:
                        os.stat(output_folder)
                    except:
                        os.mkdir(output_folder)

                    libdestino = xlwt.Workbook()
                    hojadestino = libdestino.add_sheet("Reporte")

                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=becas_pagos_' + random.randint(1,
                                                                                                           10000).__str__() + '.xls'
                    nombre = "BECAS_PAGOS_" + datetime.now().strftime('%Y%m%d_%H%M%S') + ".xls"
                    filename = os.path.join(output_folder, nombre)
                    ruta = "media/becaspregrado/" + nombre

                    hojadestino.write_merge(0, 0, 0, 15, 'UNIVERSIDAD ESTATAL DE MILAGRO', titulo)
                    hojadestino.write_merge(1, 1, 0, 15, 'DIRECCIÓN DE BIENESTAR UNIVERSITARIO', titulo2)
                    hojadestino.write_merge(2, 2, 0, 15, tituloreporte, titulo2)
                    hojadestino.write_merge(3, 3, 0, 15, 'PERIODO ' + periodo.nombre, titulo2)

                    fila = 5

                    columnas = [
                        (u"NOMBRES COMPLETOS", 10000),
                        (u"IDENTIFICACIÓN", 5000),
                        (u"FACULTAD", 5000),
                        (u"CARRERA", 5000),
                        (u"MODALIDAD", 5000),
                        (u"NIVEL", 5000),
                        (u"CORREO PERSONAL", 5000),
                        (u"CORREO UNEMI", 5000),
                        (u"TELÉFONO", 5000),
                        (u"CELULAR", 5000),
                        (u"OPERADORA", 5000),
                        (u"TIPO BECA", 8000),
                        (u"MONTO BECA", 3500),
                        (u"NÚMERO REPORTE", 3000),
                        (u"FECHA SOL.PAGO", 3000),
                        (u"ESTADO", 3000),
                        (u"FECHA PAGO", 3000)
                    ]

                    for col_num in range(len(columnas)):
                        hojadestino.write(fila, col_num, columnas[col_num][0], fuentecabecera)
                        hojadestino.col(col_num).width = columnas[col_num][1]

                    fila = 6
                    for solicitudpago in becassolicitudpago:
                        personasol = solicitudpago.asignacion.solicitud.inscripcion.persona
                        beca = solicitudpago.asignacion
                        matricula = None

                        if beca.solicitud.inscripcion.matricula_periodo_actual(periodo):
                            matricula = beca.solicitud.inscripcion.matricula_periodo_actual(periodo)[0]

                        hojadestino.write(fila, 0, personasol.nombre_completo_inverso(), fuentenormal)
                        hojadestino.write(fila, 1, personasol.identificacion(), fuentenormal)
                        hojadestino.write(fila, 2, str(matricula.nivel.coordinacion()) if matricula else '',
                                          fuentenormal)
                        hojadestino.write(fila, 3, str(matricula.inscripcion.carrera) if matricula else '',
                                          fuentenormal)
                        hojadestino.write(fila, 4, str(matricula.inscripcion.modalidad) if matricula else '',
                                          fuentenormal)
                        hojadestino.write(fila, 5, matricula.nivelmalla.nombre if matricula else '', fuentenormal)
                        hojadestino.write(fila, 6, personasol.email, fuentenormal)
                        hojadestino.write(fila, 7, personasol.emailinst, fuentenormal)
                        hojadestino.write(fila, 8, personasol.telefono_conv, fuentenormal)
                        hojadestino.write(fila, 9, personasol.telefono, fuentenormal)
                        hojadestino.write(fila, 10,
                                          personasol.get_tipocelular_display() if personasol.tipocelular else '',
                                          fuentenormal)
                        hojadestino.write(fila, 11, beca.solicitud.becatipo.nombre.upper(), fuentenormal)
                        hojadestino.write(fila, 12, beca.montobeneficio, fuentemoneda)
                        hojadestino.write(fila, 13, str(solicitudpago.solicitudpago.numerosolicitud).zfill(5),
                                          fuentenormal)
                        hojadestino.write(fila, 14, solicitudpago.solicitudpago.fecha, fuentefecha)
                        hojadestino.write(fila, 15, "PAGADO" if solicitudpago.pagado else "NO PAGADO", fuentenormal)
                        hojadestino.write(fila, 16, solicitudpago.fechapago if solicitudpago.fechapago else "",fuentefecha)
                        print(fila - 5)

                        fila += 1

                    libdestino.save(filename)
                    return JsonResponse({'result': 'ok', 'archivo': ruta})
                else:
                    return JsonResponse({'result': 'bad', 'mensaje': 'No existen registros para generar el reporte'})
            except Exception as ex:
                msg = ex.__str__()
                return JsonResponse({"result": "bad", "mensaje": u"Error al generar el reporte. [%s]" % msg})

        elif action == 'reporte_becas_acreditadas':
            try:
                estado = int(request.POST['estado'])
                if estado == 1 or estado == 2:
                    estadoacredita = (estado == 1)
                    tituloreporte = "REPORTE DE BECAS ACREDITADAS" if estadoacredita else "REPORTE DE BECAS NO ACREDITADAS"
                    becassolicitudpago = SolicitudPagoBecaDetalle.objects.filter(status=True,
                                                                                 asignacion__solicitud__periodo=periodo,
                                                                                 pagado=True,
                                                                                 acreditado=estadoacredita).distinct().order_by(
                        'asignacion__solicitud__inscripcion__persona__apellido1',
                        'asignacion__solicitud__inscripcion__persona__apellido2',
                        'asignacion__solicitud__inscripcion__persona__nombres')
                else:
                    tituloreporte = "REPORTE DE BECAS ACREDITADAS Y NO ACREDITADAS"
                    becassolicitudpago = SolicitudPagoBecaDetalle.objects.filter(status=True,
                                                                                 asignacion__solicitud__periodo=periodo,
                                                                                 pagado=True).distinct().order_by(
                        'asignacion__solicitud__inscripcion__persona__apellido1',
                        'asignacion__solicitud__inscripcion__persona__apellido2',
                        'asignacion__solicitud__inscripcion__persona__nombres')

                if becassolicitudpago:
                    fuentenormal = easyxtitulo = easyxf(
                        'font: name Times New Roman, color-index black, bold on , height 350; alignment: horiz centre')
                    titulo = easyxf(
                        'font: name Times New Roman, color-index black, bold on , height 350; alignment: horiz centre')
                    titulo2 = easyxf(
                        'font: name Times New Roman, color-index black, bold on , height 250; alignment: horiz centre')
                    fuentecabecera = easyxf(
                        'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
                    fuentenormal = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
                    fuentenormalwrap = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
                    fuentenormalneg = easyxf(
                        'font: name Verdana, color-index black, bold on , height 150; borders: left thin, right thin, top thin, bottom thin')
                    fuentenormalnegrell = easyxf(
                        'font: name Verdana, color-index black, bold on , height 150; borders: left thin, right thin, top thin, bottom thin; pattern: pattern solid, fore_colour gray25')
                    fuentenormalwrap.alignment.wrap = True
                    fuentenormalcent = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz centre')
                    fuentemoneda = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right',
                        num_format_str=' "$" #,##0.00')
                    fuentemonedaneg = easyxf(
                        'font: name Verdana, color-index black, bold on, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right; pattern: pattern solid, fore_colour gray25',
                        num_format_str=' "$" #,##0.00')
                    fuentefecha = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz center',
                        num_format_str='yyyy-mm-dd')
                    fuentenumerodecimal = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right',
                        num_format_str='#,##0.00')
                    fuentenumeroentero = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right')

                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False

                    output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media', 'becaspregrado'))
                    try:
                        os.stat(output_folder)
                    except:
                        os.mkdir(output_folder)

                    libdestino = xlwt.Workbook()
                    hojadestino = libdestino.add_sheet("Reporte")

                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=becas_acreditaciones_' + random.randint(1,
                                                                                                                    10000).__str__() + '.xls'
                    nombre = "BECAS_ACREDITACIONES_" + datetime.now().strftime('%Y%m%d_%H%M%S') + ".xls"
                    filename = os.path.join(output_folder, nombre)
                    ruta = "media/becaspregrado/" + nombre

                    hojadestino.write_merge(0, 0, 0, 15, 'UNIVERSIDAD ESTATAL DE MILAGRO', titulo)
                    hojadestino.write_merge(1, 1, 0, 15, 'DIRECCIÓN DE BIENESTAR UNIVERSITARIO', titulo2)
                    hojadestino.write_merge(2, 2, 0, 15, tituloreporte, titulo2)
                    hojadestino.write_merge(3, 3, 0, 15, 'PERIODO ' + periodo.nombre, titulo2)

                    fila = 5

                    columnas = [
                        (u"NOMBRES COMPLETOS", 10000),
                        (u"IDENTIFICACIÓN", 5000),
                        (u"FACULTAD", 5000),
                        (u"CARRERA", 5000),
                        (u"MODALIDAD", 5000),
                        (u"NIVEL", 5000),
                        (u"CORREO PERSONAL", 5000),
                        (u"CORREO UNEMI", 5000),
                        (u"TELÉFONO", 5000),
                        (u"CELULAR", 5000),
                        (u"OPERADORA", 5000),
                        (u"TIPO BECA", 8000),
                        (u"MONTO BECA", 3500),
                        (u"NÚMERO REPORTE", 3000),
                        (u"FECHA SOL.PAGO", 3000),
                        (u"FECHA PAGO", 3000),
                        (u"ESTADO", 3500),
                        (u"FECHA ACREDITACIÓN", 3000)
                    ]

                    for col_num in range(len(columnas)):
                        hojadestino.write(fila, col_num, columnas[col_num][0], fuentecabecera)
                        hojadestino.col(col_num).width = columnas[col_num][1]

                    fila = 6
                    for solicitudpago in becassolicitudpago:
                        personasol = solicitudpago.asignacion.solicitud.inscripcion.persona
                        beca = solicitudpago.asignacion
                        matricula = None

                        if beca.solicitud.inscripcion.matricula_periodo_actual(periodo):
                            matricula = beca.solicitud.inscripcion.matricula_periodo_actual(periodo)[0]

                        hojadestino.write(fila, 0, personasol.nombre_completo_inverso(), fuentenormal)
                        hojadestino.write(fila, 1, personasol.identificacion(), fuentenormal)
                        hojadestino.write(fila, 2, str(matricula.nivel.coordinacion()) if matricula else '',
                                          fuentenormal)
                        hojadestino.write(fila, 3, str(matricula.inscripcion.carrera) if matricula else '',
                                          fuentenormal)
                        hojadestino.write(fila, 4, str(matricula.inscripcion.modalidad) if matricula else '',
                                          fuentenormal)
                        hojadestino.write(fila, 5, matricula.nivelmalla.nombre if matricula else '', fuentenormal)
                        hojadestino.write(fila, 6, personasol.email, fuentenormal)
                        hojadestino.write(fila, 7, personasol.emailinst, fuentenormal)
                        hojadestino.write(fila, 8, personasol.telefono_conv, fuentenormal)
                        hojadestino.write(fila, 9, personasol.telefono, fuentenormal)
                        hojadestino.write(fila, 10,
                                          personasol.get_tipocelular_display() if personasol.tipocelular else '',
                                          fuentenormal)
                        hojadestino.write(fila, 11, beca.solicitud.becatipo.nombre.upper(), fuentenormal)
                        hojadestino.write(fila, 12, beca.montobeneficio, fuentemoneda)
                        hojadestino.write(fila, 13, str(solicitudpago.solicitudpago.numerosolicitud).zfill(5),
                                          fuentenormal)
                        hojadestino.write(fila, 14, solicitudpago.solicitudpago.fecha, fuentefecha)
                        hojadestino.write(fila, 15, solicitudpago.fechapago, fuentefecha)
                        hojadestino.write(fila, 16, "ACREDITADO" if solicitudpago.acreditado else "NO ACREDITADO",
                                          fuentenormal)
                        hojadestino.write(fila, 17, solicitudpago.fechaacredita if solicitudpago.fechaacredita else "",
                                          fuentefecha)

                        print(fila - 5)

                        fila += 1

                    libdestino.save(filename)
                    return JsonResponse({'result': 'ok', 'archivo': ruta})
                else:
                    return JsonResponse({'result': 'bad', 'mensaje': 'No existen registros para generar el reporte'})
            except Exception as ex:
                msg = ex.__str__()
                return JsonResponse({"result": "bad", "mensaje": u"Error al generar el reporte. [%s]" % msg})

        elif action == 'adddetallerequisitobeca':
            try:
                form = DetalleRequisitoBecaForm(request.POST)
                registro = DetalleRequisitoBeca()
                if not form.is_valid():
                    # [(k, v[0]) for k, v in f.errors.items()]
                    for k, v in form.errors.items():
                        raise NameError(f'{k}:{v[0]}')
                    
                funcionejecutar = form.cleaned_data['funcionejecutar']
                registro.requisitobeca = form.cleaned_data['requisitobeca']
                registro.numero = form.cleaned_data['numero']
                registro.obligatorio = form.cleaned_data['obligatorio']
                registro.visible = form.cleaned_data['visible']
                registro.funcionejecutar = funcionejecutar if funcionejecutar else None
                registro.save(request)
                log(u'Asiciono Detalle  Requisito Beca: %s' % registro, request, "add")
                return JsonResponse({"result": True, "mensaje": u"Registro Guardado Correctamente"})
            except Exception as ex:
                msg = str(ex)
                transaction.set_rollback(True)
                return JsonResponse({"result": False, "mensaje": u"Error %s" %msg})

        elif action == 'editdetallerequisitobeca':
            try:
                registro = DetalleRequisitoBeca.objects.get(pk=int(encrypt(request.POST['id'])))
                form = DetalleRequisitoBecaForm(request.POST)

                if not form.is_valid():
                    # [(k, v[0]) for k, v in f.errors.items()]
                    for k, v in form.errors.items():
                        raise NameError(f'{k}:{v[0]}')
                funcionejecutar = form.cleaned_data['funcionejecutar']
                registro.requisitobeca = form.cleaned_data['requisitobeca']
                registro.numero = form.cleaned_data['numero']
                registro.obligatorio = form.cleaned_data['obligatorio']
                registro.visible = form.cleaned_data['visible']
                registro.funcionejecutar = funcionejecutar if funcionejecutar else None
                registro.save(request)
                log(u'Edito Detalle  Requisito Beca: %s' % registro, request, "edit")
                return JsonResponse({"result": True, "mensaje": u"Registro Editado Correctamente"})
            except Exception as ex:
                msg = str(ex)
                transaction.set_rollback(True)
                return JsonResponse({"result": False, "mensaje": u"Error %s" % msg})

        elif action == 'deletedetallerequisitobeca':
            try:
                registro = DetalleRequisitoBeca.objects.get(pk=int(encrypt(request.POST['id'])))
                if not registro.esta_en_uso():
                    registro.status = False
                    registro.save(request)
                    log(u'Elimino Detalle  Requisito Beca: %s' % registro, request, "del")
                    return JsonResponse({"result": True, "mensaje": u"Registro Eliminado Correctamente"})
                else:
                    return JsonResponse({"result": False, "mensaje": u"No se puede eliminar el registro por que ya se encuentra utilizado."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": False, "mensaje": u"Error al eliminar los datos."})

        elif action == 'cambiarestadorequisitobeca':
            try:
                registro = PreInscripcionBecaRequisito.objects.get(pk=int(encrypt(request.POST['id'])))
                registro.cumplerequisito = True if request.POST.get('cumplerequisito') == 'on' else False
                registro.save(request)
                estado = """ 'Cumple' """ if registro.cumplerequisito else """ 'No Cumple' """
                observacion = request.POST['observacion']
                historial = HistorialPreInscripcionBecaRequisito(preinscripcionbecarequisito=registro,
                                                                 observacion=observacion,
                                                                 persona=persona,
                                                                 cumplerequisito=registro.cumplerequisito)
                historial.save(request)
                log(u'Cambio estado  de requisito beca: %s el estado  es %s' %(registro, estado), request, "edit")
                generar_solicitud = registro.preinscripcion.cumple_requisitos()
                registro_json = model_to_dict(registro, exclude=['archivo'])
                return JsonResponse({"result": True, "mensaje": u"Registro actulizado correctamente", "registro": registro_json, 'generar_solicitud':generar_solicitud})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": False, "mensaje": u"Error al eliminar los datos."})

        elif action == 'addupdatepreinscripcionbecarequisitos':
            try:
                GenerateBackground(request=request, data=data, periodoactual=periodo, periodoanterior=anterior).start()
                # name_document = 'estudiantes_por_anio'
                # noti = Notificacion(cuerpo='Generación requisitos de becas',
                #                     titulo=name_document, destinatario=persona,
                #                     url='',
                #                     prioridad=1, app_label='SGA',
                #                     fecha_hora_visible=datetime.now() + timedelta(days=1), tipo=2,
                #                     en_proceso=True)
                # noti.save(request)
                return JsonResponse({"result": True, "mensaje": u"Se ha procedido a ejecutar el proceso, en cuanto este se procedera a notificar"})
            except Exception as ex:
                msg = str(ex)
                return JsonResponse({"result": False, "mensaje": u"Error al ejecuta acción: %s"%msg})

        elif action == 'importar_archivopreinscriptos_becas':
            try:
                data['title'] = u'Importar archivo preinscritos'
                data['form'] = form = ImportarPreinscritosBecaForm(request.POST, request.FILES)
                newfile = None
                nombre_subido = None
                if 'archivo' in request.FILES:
                    newfile = request.FILES['archivo']

                    if newfile:
                        nombre_subido = newfile._name
                        if newfile.size > 3145728:
                            return NameError(u"Error, archivo mayor a 3 Mb.")
                        else:
                            newfilesd = newfile._name
                            ext = newfilesd[newfilesd.rfind("."):]
                        if ext in ['.xlsx', '.XLSX']: #['.xls', '.XLS', '.xlsx', '.XLSX']
                            newfile._name = generar_nombre("plantill_preinscritos_requisitos_beca", newfile._name)
                        else:
                            raise NameError(u"Solo archivo con extención (.xlsx, .xls)") #(.xlsx, .xls)

                if not form.is_valid():
                    # [(k, v[0]) for k, v in f.errors.items()]
                    for k, v in form.errors.items():
                        raise NameError(f'{k}:{v[0]}')

                wb = load_workbook(filename=newfile, read_only=True)
                sheet = wb[wb.sheetnames[0]]
                ids = {}
                idp = sheet.cell(row=2, column=1).value
                if not idp:
                    raise NameError(u"Plantilla de Archivo es Incorrecto. revise el formato del archivo")
                if not isinstance(idp, int):
                    raise NameError(u"El id de celda verificación   no es un número")

                id = int(encrypt(request.POST['id']))
                becatipo = BecaTipo.objects.get(pk=id)
                configuracion = becatipo.becatipoconfiguracion_set.filter(becaperiodo__periodo=periodo).first()
                mensaje_error = ''
                if not configuracion:
                    raise NameError('No existe configuración de requisitos para cargar archivo')
                else:
                    if not configuracion.requisitosbecas:
                        raise NameError('Existe configuración pero no contiene requisitos')
                requisitos = configuracion.requisitosbecas
                if requisitos is None:
                    raise NameError('No existe configuración de requisitos en la configuración')

                for col in range(2, sheet.max_column, 2):
                    value = sheet.cell(row=1, column=col).value.split()[0]
                    value2 = sheet.cell(row=1, column=col + 1).value.split()[0]
                    idreq = ''.join(filter(str.isalnum, value))
                    idob = ''.join(filter(str.isalnum, value2))
                    if idreq == idob:
                        if not idreq in list(ids.values()):
                            validareq = requisitos.filter(id=int(idreq)).values('id')
                            if validareq:
                                ids[col] = int(idreq)
                            else:
                                raise NameError(f'El id {idreq}  no se encuentra en los requisitos configurados en columna {col}')
                        else:
                            raise NameError(f'Los id de los requsitos deben ser únicos  en columna {col}')
                    else:
                        raise NameError(f'Los id de los requsitos no  son iguales a los de observación  en columna {col}')
                data['ids'] = ids
                requisito_archivo = PreInscripcionBecaRequisitoArchivo.objects.filter(becatipoconfiguracion=configuracion, persona=persona, estado=1)
                # with transaction.atomic(using='default'):
                #     try:
                if not requisito_archivo:
                    requisito_archivo = PreInscripcionBecaRequisitoArchivo(
                        becatipoconfiguracion=configuracion,
                        persona=persona,
                        nombre=nombre_subido,
                        archivo=newfile,
                        fecha=datetime.now().date(),
                        hora=datetime.now().time(),
                    )
                    requisito_archivo.save(request)
                    log(u'Importando Archivo  de requisitos becas: %s  en estado %s' %(requisito_archivo, requisito_archivo.get_estado_display()), request, "add")
                else:
                    raise NameError(f'Se esta procesando un archivo de la configuración tipo beca {configuracion}')
                    # except Exception as ex:
                    #     msg = str(ex)
                    #     transaction.set_rollback(True, 'default')
                    #     return JsonResponse({"result": False, "mensaje": u"Error al validar el archivo: %s" % msg})

                name_document = f'Importando Archivo de requsitos tipo beca {becatipo}'
                noti = Notificacion(cuerpo=f'Importacion  de requisitos de tipo beca {becatipo} en el archivo {nombre_subido}',
                                    titulo=name_document, destinatario=persona,
                                    url='/adm_becas',
                                    prioridad=1, app_label='SGA',
                                    fecha_hora_visible=datetime.now() + timedelta(days=1), tipo=2,
                                    en_proceso=True)
                noti.save(request)
                data['requisito_archivo'] = requisito_archivo
                GenerateBackground(request=request, data=data, periodoactual=periodo, periodoanterior=anterior, noti=noti).start()
                return JsonResponse({"result": True, "mensaje": u"Se ha procedido a ejecutar el proceso, en cuanto este se procedera a notificar"})
            except Exception as ex:
                msg = str(ex)
                return JsonResponse({"result": False, "mensaje": u"Error al validar el archivo: %s" % msg})

        elif action == 'importar_archivoexcelbecarios':
            try:
                data['form'] = form = ImportarArchivoBecariosForm(request.POST, request.FILES)
                newfile = None
                nombre_subido = None
                if 'archivo' in request.FILES:
                    newfile = request.FILES['archivo']

                    if newfile:
                        nombre_subido = newfile._name
                        if newfile.size > 3145728:
                            return NameError(u"Error, archivo mayor a 3 Mb.")
                        else:
                            newfilesd = newfile._name
                            ext = newfilesd[newfilesd.rfind("."):]
                        if ext in ['.xlsx', '.XLSX']: #['.xls', '.XLS', '.xlsx', '.XLSX']
                            newfile._name = generar_nombre("plantill_preinscritos_requisitos_beca", newfile._name)
                        else:
                            raise NameError(u"Solo archivo con extención (.xlsx)") #(.xlsx, .xls)

                if not form.is_valid():
                    for k, v in form.errors.items():
                        raise NameError(f'{k}:{v[0]}')

                wb = load_workbook(filename=newfile, read_only=True)
                sheet = wb[wb.sheetnames[0]]
                ids = {}
                idp = sheet.cell(row=2, column=1).value
                if not idp:
                    raise NameError(u"Celda de verificación se encuentra vacía")
                if not isinstance(idp, int):
                    raise NameError(u"El id de verificación  no es un número") #(.xlsx, .xls)

                preinscripcionbeca_verificacion = PreInscripcionBeca.objects.filter(id=idp).first()
                if preinscripcionbeca_verificacion is None:
                    raise NameError(u"No existe el id(%s) de verificación preinscripción beca"%idp)
                idbt = int(encrypt(request.POST['idbt']))
                becatipo = BecaTipo.objects.filter(id=idbt).first()
                becatipo_verificacion = preinscripcionbeca_verificacion.becatipo
                if becatipo_verificacion.id != becatipo.id:
                    raise NameError(u"El archivo en su celda de verificación contiene registro de Tipo beca %s cuando debe ser %s"%(becatipo_verificacion, becatipo))
                data['wb'] = wb
                data['newfile'] = newfile
                data['becatipo'] = becatipo

                archivobecarios = ArchivoBecarios(
                    persona=persona,
                    nombre='Importación de becarios',
                    archivo=newfile,
                    fecha=datetime.now().date(),
                    hora=datetime.now().time())
                archivobecarios.save(request)
                data['archivobecarios'] = archivobecarios
                name_document = 'Importando archivo excel de becados'
                noti = Notificacion(cuerpo='Importación  de Archivo Excel Becados',
                                    titulo=name_document, destinatario=persona,
                                    url='/adm_becas',
                                    prioridad=1, app_label='SGA',
                                    fecha_hora_visible=datetime.now() + timedelta(days=1), tipo=2,
                                    en_proceso=True)
                noti.save(request)
                #log(u'Importando Archivo excel de becados:cargo archivo  %s', request.user, "add")
                GenerateBackground(request=request, data=data, periodoactual=periodo, periodoanterior=anterior, noti=noti).start()
                return JsonResponse({"result": True, "mensaje": u"Se ha procedido a ejecutar el proceso, en cuanto este se procedera a notificar"})
            except Exception as ex:
                msg = str(ex)
                return JsonResponse({"result": False, "mensaje": u"Error al validar el archivo: %s" % msg})

        elif action == 'validar_data':
            try:
                data['title'] = u'Importar archivo preinscritos'
                data['form'] = form = ImportarPreinscritosBecaForm(request.POST, request.FILES)
                newfile = None
                if 'archivo' in request.FILES:
                    newfile = request.FILES['archivo']
                    if newfile:
                        # if newfile.size > 3145728:
                        #     return NameError(u"Error, archivo mayor a 3 Mb.")
                        # else:
                        newfilesd = newfile._name
                        ext = newfilesd[newfilesd.rfind("."):]
                        if ext  in ['.xls', '.XLS', '.xlsx', '.XLSX'] :
                            newfile._name = generar_nombre("formulario_", newfile._name)
                        else:
                            raise NameError(u"Solo archivo con extención. pdf")


                data['action'] = action
                data['user'] = request.user
                template = get_template("adm_becas/modal/tableexportarpreinscritosbeca.html")
                return JsonResponse({"result": True, 'table': template.render(data)})
            except Exception as ex:
                msg = str(ex)
                transaction.set_rollback(True)
                return JsonResponse({"result": False, "mensaje": u"Error al validar el archivo: %s" % msg})

        elif action == 'uploadarchivorequisito':
            try:
                form = SubirArchivoRequisitoBecaForm(request.POST, request.FILES)
                if not form.is_valid():
                    # [(k, v[0]) for k, v in f.errors.items()]
                    for k, v in form.errors.items():
                        raise NameError(f'{k}:{v[0]}')
                id = int(encrypt(request.POST['id']))
                registro = PreInscripcionBecaRequisito.objects.get(pk=id)
                newfile = None
                if 'archivo' in request.FILES:
                    newfile = request.FILES['archivo']
                    if newfile:
                        # if newfile.size > 3145728:
                        #     return NameError(u"Error, archivo mayor a 3 Mb.")
                        # else:
                        inscripcion = registro.preinscripcion.inscripcion
                        username = inscripcion.persona.usuario.username
                        periodo_id = registro.preinscripcion.periodo_id
                        # namerequisito = registro.detallerequisitobeca.requisito.nombre.replace(' ','_')


                        newfilesd = newfile._name
                        ext = newfilesd[newfilesd.rfind("."):]
                        if ext in ['.pdf', '.PDF']:
                            namefile = f'requisitobeca_{inscripcion.id}_{registro.detallerequisitobeca.id}_'
                            filename = f'{periodo_id}/{username}/{namefile}'
                            newfile._name = generar_nombre(filename , newfile._name)
                        else:
                            raise NameError(u"Solo archivo con extención. pdf")

                registro.archivo = newfile
                registro.save(request)
                log(u'Cargo Archivo de  requisito beca: %s' % registro, request, "edit")
                historial = HistorialPreInscripcionBecaRequisito(preinscripcionbecarequisito=registro,
                                                                 observacion='Cargo Archivo',
                                                                 persona=persona,
                                                                 cumplerequisito=registro.cumplerequisito)
                historial.save(request)
                data_archivo = {
                    'name': newfile._name,
                    'url': registro.archivo.url,
                    'buttonaction': f'#btn-download-archivorequisito{registro.id}'}
                return JsonResponse({"result": "ok", 'mensaje': u'Cargo correctamente. el archivo %s' % (registro), 'archivo':data_archivo})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error: %s" % str(ex)})

        elif action == 'deletepreinscripcionbeca':
            try:
                registro = PreInscripcionBeca.objects.get(pk=request.POST['id'])
                registro.delete()
                log(u'Eliminó preinscritos de becas: %s' % registro, request, "del")
                return JsonResponse({"result": "ok", 'mensaje': u'Registro eliminado correctamente.'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        elif action == 'saveRequerement':
            try:
                id = int(encrypt(request.POST['id'])) if 'id' in request.POST and request.POST['id'] and int(encrypt(request.POST['id'])) != 0 else None
                typeForm = 'edit' if id else 'new'
                if typeForm == 'new':
                    eRequisitobeca = RequisitoBeca(
                        nombre=request.POST['nombre']
                    )
                    eRequisitobeca.save(request)
                    log(u'Agrego requisito beca: %s' % eRequisitobeca, request, "add")
                else:
                    eRequisitobeca = RequisitoBeca.objects.get(pk=id)
                    eRequisitobeca.nombre = request.POST['nombre']
                    eRequisitobeca.save(request)
                    log(u'Edito requisito beca: %s' % eRequisitobeca, request, "edit")

                return JsonResponse({"result": "ok", "mensaje": u"Se guardo correctamente el requisito"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar requisito. %s" % ex.__str__()})

        elif action == 'deleteRequerement':
            try:
                id = int(encrypt(request.POST['id'])) if 'id' in request.POST and request.POST['id'] and int(encrypt(request.POST['id'])) != 0 else None
                eRequisitobeca = RequisitoBeca.objects.get(pk=id)
                if eRequisitobeca.esta_en_uso:
                        raise NameError('Requisito esta en uso')
                eRequisitobeca.status = False
                eRequisitobeca.save(request)
                log(u'Elimino requisito beca: %s' % eRequisitobeca, request, "edit")
                return JsonResponse({"result": "ok", "mensaje": u"Se elimino correctamente el requisito"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar requisito. %s" % ex.__str__()})

        elif action == 'adddocumentobecas':
            try:
                form = DocumentoBecaTipoConfiguracionForm(request.POST)
                registro = DocumentoBecaTipoConfiguracion()
                if not form.is_valid():
                    # [(k, v[0]) for k, v in f.errors.items()]
                    for k, v in form.errors.items():
                        raise NameError(f'{k}:{v[0]}')

                funcionejecutar = form.cleaned_data['funcionejecutar']
                registro.tipo = form.cleaned_data['tipo']
                registro.numero = form.cleaned_data['numero']
                registro.obligatorio = form.cleaned_data['obligatorio']
                registro.visible = form.cleaned_data['visible']
                registro.funcionejecutar = funcionejecutar if funcionejecutar else None
                registro.save(request)
                log(u'Asiciono Detalle  Requisito Beca: %s' % registro, request, "add")
                return JsonResponse({"result": True, "mensaje": u"Registro Guardado Correctamente"})
            except Exception as ex:
                msg = str(ex)
                transaction.set_rollback(True)
                return JsonResponse({"result": False, "mensaje": u"Error %s" % msg})

        elif action == 'editdocumentobecas':
            try:
                registro = DocumentoBecaTipoConfiguracion.objects.get(pk=int(encrypt(request.POST['id'])))
                form = DocumentoBecaTipoConfiguracionForm(request.POST)

                if not form.is_valid():
                    # [(k, v[0]) for k, v in f.errors.items()]
                    for k, v in form.errors.items():
                        raise NameError(f'{k}:{v[0]}')
                funcionejecutar = form.cleaned_data['funcionejecutar']
                registro.tipo = form.cleaned_data['tipo']
                registro.numero = form.cleaned_data['numero']
                registro.obligatorio = form.cleaned_data['obligatorio']
                registro.visible = form.cleaned_data['visible']
                registro.funcionejecutar = funcionejecutar if funcionejecutar else None
                registro.save(request)
                log(u'Edito Detalle  Requisito Beca: %s' % registro, request, "edit")
                return JsonResponse({"result": True, "mensaje": u"Registro Editado Correctamente"})
            except Exception as ex:
                msg = str(ex)
                transaction.set_rollback(True)
                return JsonResponse({"result": False, "mensaje": u"Error %s" % msg})

        elif action == 'deletedocumentobecas':
            try:
                registro = DetalleRequisitoBeca.objects.get(pk=int(encrypt(request.POST['id'])))
                if not registro.esta_en_uso():
                    registro.status = False
                    registro.save(request)
                    log(u'Elimino Detalle  Requisito Beca: %s' % registro, request, "del")
                    return JsonResponse({"result": True, "mensaje": u"Registro Eliminado Correctamente"})
                else:
                    return JsonResponse({"result": False, "mensaje": u"No se puede eliminar el registro por que ya se encuentra utilizado."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": False, "mensaje": u"Error al eliminar los datos."})

        elif action == 'corregir_actascompromisos_becas':
            try:
                noti = Notificacion(cuerpo=f'Proceso de correción  de  actas de compromisos becas período académico {periodo}',
                                    titulo=f'Ejecución en proceso de actas compromiso {periodo}', destinatario=persona,
                                    url='/adm_becas',
                                    prioridad=1, app_label='SGA',
                                    fecha_hora_visible=datetime.now() + timedelta(days=1), tipo=2,
                                    en_proceso=True)
                noti.save(request)
                GenerateBackground(request=request, data=data, periodoactual=periodo, periodoanterior=anterior, noti=noti).start()
                return JsonResponse({"result": True, "mensaje": u"Se ha procedido a ejecutar el proceso, en cuanto este se procedera a notificar"})
            except Exception as ex:
                msg = str(ex)
                return JsonResponse({"result": False, "mensaje": u"Error %s" % msg})

        elif action == 'loadDataTableBecados':
            try:
                txt_filter = request.POST['sSearch'] if request.POST['sSearch'] else ''
                limit = int(request.POST['iDisplayLength']) if request.POST['iDisplayLength'] else 25
                offset = int(request.POST['iDisplayStart']) if request.POST['iDisplayStart'] else 0
                tCount = 0
                becatipo = request.POST['becatipo']
                estadodocumento = request.POST['estadodocumento']
                asignados = BecaAsignacion.objects.filter(status=True,
                                                          solicitud__inscripcion__persona__personadocumentopersonal__status=True,
                                                          solicitud__inscripcion__persona__personadocumentopersonal__cedula__isnull=False,
                                                          solicitud__inscripcion__persona__personadocumentopersonal__estadocedula__isnull=False,
                                                          solicitudpagobecadetalle__isnull=True,
                                                          solicitud__periodo=periodo,
                                                          solicitud__becatipo_id__isnull=False).exclude(solicitud__inscripcion__persona__personadocumentopersonal__estadocedula=2).distinct() \
                    .order_by('-fecha_creacion', 'solicitud__inscripcion__persona__apellido1',
                              'solicitud__inscripcion__persona__apellido2', 'solicitud__inscripcion__persona__nombres')
                if becatipo:
                    becatipo = int(becatipo)
                    asignados = asignados.filter(solicitud__becatipo_id=becatipo).order_by('-fecha_creacion', 'solicitud__inscripcion__persona__apellido1',
                              'solicitud__inscripcion__persona__apellido2', 'solicitud__inscripcion__persona__nombres')
                if estadodocumento:
                    estadodocumento = int(estadodocumento)
                    asignados = asignados.filter(solicitud__inscripcion__persona__personadocumentopersonal__estadocedula=estadodocumento).distinct().order_by('-fecha_creacion', 'solicitud__inscripcion__persona__apellido1',
                              'solicitud__inscripcion__persona__apellido2', 'solicitud__inscripcion__persona__nombres')

                if txt_filter:
                    search = txt_filter.strip()
                    ss = search.split(' ')
                    if len(ss) == 1:
                        asignados = asignados.filter(Q(solicitud__inscripcion__persona__cedula__icontains=search) |
                                                     Q(solicitud__inscripcion__persona__nombres__icontains=search) |
                                                     Q(solicitud__inscripcion__persona__apellido1__icontains=search) |
                                                     Q(solicitud__inscripcion__persona__apellido2__icontains=search) |
                                                     Q(solicitud__becatipo__nombre__icontains=search)) \
                            .order_by('-fecha_creacion',
                                      'solicitud__inscripcion__persona__apellido1',
                                      'solicitud__inscripcion__persona__apellido2',
                                      'solicitud__inscripcion__persona__nombres')
                    else:
                        asignados = asignados.filter(Q(solicitud__inscripcion__persona__nombres__icontains=ss[0]) |
                                                     Q(solicitud__inscripcion__persona__apellido1__icontains=ss[0]) |
                                                     Q(solicitud__inscripcion__persona__apellido2__icontains=ss[0]) |
                                                     Q(solicitud__becatipo__nombre__icontains=ss[0]) |
                                                     Q(solicitud__inscripcion__persona__nombres__icontains=ss[1]) |
                                                     Q(solicitud__inscripcion__persona__apellido1__icontains=ss[1]) |
                                                     Q(solicitud__inscripcion__persona__apellido2__icontains=ss[1]) |
                                                     Q(solicitud__becatipo__nombre__icontains=ss[1])) \
                            .order_by('-fecha_creacion',
                                      'solicitud__inscripcion__persona__apellido1',
                                      'solicitud__inscripcion__persona__apellido2',
                                      'solicitud__inscripcion__persona__nombres')
                tCount = asignados.count()
                if offset == 0:
                    rows = asignados[offset:limit]
                else:
                    rows = asignados[offset:offset + limit]
                aaData = []
                for row in rows:
                    solicitante = row.solicitud.inscripcion.persona
                    documentos = solicitante.documentos_personales()
                    documento = {}
                    urlcedula = ''
                    if documentos:
                        urlcedula = documentos.cedula.url if  documentos.cedula else ''
                        documento = {
                            'url': urlcedula,
                            'estado': documentos.estadocedula,
                            'estado_diplay': documentos.get_estadocedula_display()
                        }
                    aaData.append([row.id,
                                   {'nombres': row.solicitud.inscripcion.__str__(),
                                    'becatipo': row.solicitud.becatipo.__str__(),
                                    'tipo': row.get_tipo_display(),
                                    'fecha': row.fecha_creacion.strftime('%Y-%m-%d %H:%M:%S'),
                                    'documento':solicitante.documento(),
                                    'tipo_documento':solicitante.tipo_documento(),
                                    'carrera':row.solicitud.inscripcion.carrera.__str__(),
                                    },
                                   {
                                       'telefonos': ','.join(solicitante.lista_telefonos()),
                                       'direccion': solicitante.direccion_completa(),
                                       'emails': solicitante.lista_emails(),
                                   },
                                    documento,
                                    urlcedula])
                return JsonResponse({"result": "ok", "data": aaData, "iTotalRecords": tCount, "iTotalDisplayRecords": tCount})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al cargar los datos. %s" % ex.__str__(), "data": [], "iTotalRecords": 0, "iTotalDisplayRecords": 0})

        elif action == 'savelBecadosMasivoDocumentos':
            try:
                import time
                becados = request.POST.get('becados')
                observacion = request.POST.get('observaciones', '')
                estadodocumentoasignar = request.POST.get('estadodocumentoasignar')
                becados = json.loads(becados)
                becaperiodo = BecaPeriodo.objects.filter(periodo=periodo, vigente=True).first()
                # if not becaperiodo.puede_revisar_validar_documentos():
                #     return JsonResponse({"result": "bad", "mensaje":"Se completo el límite de becados del período %s"%(becaperiodo)})
                if not becados:
                    raise NameError('No selecciono ningun becado')
                
                for beca_id in becados:
                    beca = BecaAsignacion.objects.get(pk=int(beca_id))
                    documentos = beca.solicitud.inscripcion.persona.documentos_personales()
                    documentos.estadocedula = int(estadodocumentoasignar)
                    documentos.observacion = request.POST['observacion'].strip().upper() if 'observacion' in request.POST else ''
                    documentos.save(request)
                    lista_correos = beca.solicitud.inscripcion.persona.lista_emails_envio(),
                    if DEBUG:
                        lista_correos = ['atorrese@unemi.edu.ec', ]
                    if documentos.estadocedula == 3 :
                        tituloemail = "Novedades con Documentación"
                        mensaje = ""
                        if documentos.estadocedula == 3:
                            mensaje = " se presentaron novedades con la revisión del archivo de su cédula de identidad"


                        observaciones = documentos.observacion

                        recorrido = BecaSolicitudRecorrido(solicitud=beca.solicitud,
                                                           observacion=observaciones,
                                                           estado=12,
                                                           fecha=datetime.now().date())
                        recorrido.save(request)

                        send_html_mail(tituloemail,
                                       "emails/notificarrevisiondocumento.html",
                                       {'sistema': u'SGA - UNEMI',
                                        'mensaje': mensaje,
                                        'fecha': datetime.now().date(),
                                        'hora': datetime.now().time(),
                                        'observaciones': observaciones,
                                        'saludo': 'Estimada' if beca.solicitud.inscripcion.persona.sexo_id == 1 else 'Estimado',
                                        'estudiante': beca.solicitud.inscripcion.persona.nombre_completo_inverso(),
                                        'autoridad2': '',
                                        't': miinstitucion()
                                        },
                                       lista_correos,
                                       [],
                                       cuenta=variable_valor('CUENTAS_CORREOS')[0]
                                       )
                    else:
                        recorrido = BecaSolicitudRecorrido(solicitud=beca.solicitud,
                                                           observacion="DOCUMENTACIÓN VALIDADA",
                                                           estado=11,
                                                           fecha=datetime.now().date())
                        recorrido.save(request)

                        tituloemail = "Beca - Documentación Validada"

                        if beca.solicitud.becatipo.id == 23:
                            mensaje = "los documentos correspondientes a la cédula a sido Validada"
                            #observaciones = "Usted deberá acceder al Módulo de Beca Estudiantil, descargar el contrato de ayuda económica, imprimirlo, recoger las firmas correspondientes y subirlo al sistema en formato PDF."
                        else:
                            if beca.solicitud.becatipo.id == 16:
                                mensaje = "los documentos correspondientes al estudiante a sido Validados"
                            else:
                                mensaje = "los documentos correspondientes a la cédula a sido Validada"

                            #observaciones = f"Usted deberá acceder al Módulo de Beca Estudiantil, descargar el contrato de beca (disponible el {leer_fecha_hora_letra_formato_fecha_hora(str(fechainicioimprimircontrato.date()), str(fechainicioimprimircontrato.time()), True)}), imprimirlo, recoger las firmas correspondientes y subirlo al sistema en formato PDF."

                        send_html_mail(tituloemail,
                                       "emails/notificarrevisiondocumento.html",
                                       {'sistema': u'SGA - UNEMI',
                                        'mensaje': mensaje,
                                        'fecha': datetime.now().date(),
                                        'hora': datetime.now().time(),
                                        'observaciones': '',
                                        'saludo': 'Estimada' if beca.solicitud.inscripcion.persona.sexo_id == 1 else 'Estimado',
                                        'estudiante': beca.solicitud.inscripcion.persona.nombre_completo_inverso(),
                                        'autoridad2': '',
                                        't': miinstitucion()
                                        },
                                       lista_correos,
                                       [],
                                       cuenta=variable_valor('CUENTAS_CORREOS')[0]
                                       )

                    log(u'Revisó documentación cédula de la beca: %s  - %s' % (persona, beca), request, "edit")
                    time.sleep(2)
                return JsonResponse({"result": True, "mensaje": u"Se verificaron correctamente los becados"})
            except Exception as ex:
                transaction.set_rollback(True)
                msg = str(ex)
                return JsonResponse({"result": False, "mensaje": u"Error %s" % msg})

        elif action == 'loadDataTablePreDataBecas':
                try:
                    txt_filter = request.POST['sSearch'] if request.POST['sSearch'] else ''
                    limit = int(request.POST['iDisplayLength']) if request.POST['iDisplayLength'] else 25
                    offset = int(request.POST['iDisplayStart']) if request.POST['iDisplayStart'] else 0
                    tCount = 0
                    becatipo_id = request.POST.get('bt')
                    beca_seleccionado = request.POST.get('bests')

                    filtros = Q(status=True, periodo_id=periodo.id)

                    if becatipo_id:
                        filtros &= Q(status=True, becatipo_id=int(becatipo_id))
                    if beca_seleccionado:
                        filtros &= Q(seleccionado=beca_seleccionado=="1")

                    if txt_filter:
                        search = txt_filter.strip()
                        ss = search.split(' ')
                        if len(ss) == 1:
                            filtros &= Q(Q(inscripcion__persona__cedula__icontains=search) |
                                         Q(inscripcion__persona__pasaporte__icontains=search) |
                                         Q(inscripcion__persona__ruc__icontains=search) |
                                         Q(inscripcion__persona__nombres__icontains=search) |
                                         Q(inscripcion__persona__apellido1__icontains=search) |
                                         Q(inscripcion__persona__apellido2__icontains=search)
                                        )
                        else:
                            filtros &= Q(Q(inscripcion__persona__nombres__icontains=search) |
                                         Q(inscripcion__persona__apellido1__icontains=ss[0]))

                    matriculados = PreInscripcionBeca.objects.filter(filtros).order_by("becatipo__nombre", "orden").distinct()
                    tCount = matriculados.count()
                    if offset == 0:
                        rows = matriculados[offset:limit]
                    else:
                        rows = matriculados[offset:offset + limit]
                    aaData = []
                    url_path = request.build_absolute_uri('/')[:-1].strip("/")
                    for row in rows:
                        matriculaactual = row.inscripcion.matricula_periodo(periodo)
                        matriculaanterior = row.inscripcion.matricula_periodo(anterior)
                        jsonmatriculaactual = {}
                        jsonmatriculaanterior = {}
                        foto_url = f"static/images/iconos/{'mujer.png' if row.inscripcion.persona.sexo_id == 1 else 'hombre.png' }"
                        foto = row.inscripcion.persona.foto()
                        foto_url = f"{url_path}{foto.foto.url if foto else foto_url}"

                        if matriculaactual is not None:
                            gruposocioeconomio = matriculaactual.matricula_gruposocioeconomico()
                            jsonmatriculaactual = {
                                'id': matriculaactual.id,
                                'tipo_matricula': gruposocioeconomio.tipomatricula,
                                'tipo_matricula_texto': gruposocioeconomio.get_tipomatricula_display(),
                                'display': f'{matriculaactual.nivelmalla.__str__()} - {matriculaactual.paralelo.nombre if matriculaactual.paralelo else ""}',
                            }
                        if matriculaanterior is not None:
                            gruposocioeconomio = matriculaanterior.matricula_gruposocioeconomico()
                            jsonmatriculaanterior = {
                                'id': matriculaanterior.id,
                                'tipo_matricula': gruposocioeconomio.tipomatricula,
                                'tipo_matricula_texto': gruposocioeconomio.get_tipomatricula_display(),
                                'display': f'{matriculaanterior.nivelmalla.__str__() } - {matriculaanterior.paralelo.nombre if matriculaanterior.paralelo else ""}',
                            }
                        aaData.append([{'nombres': row.inscripcion.persona.__str__(),
                                        'becatipo': row.becatipo.__str__() if row.becatipo else '',
                                        'becatipo_nombrecorto': row.becatipo.nombrecorto.upper() if row.becatipo else '',
                                        'fecha': row.fecha_creacion.strftime('%Y-%m-%d %H:%M:%S'),
                                        'documento': row.inscripcion.persona.documento(),
                                        'tipo_documento': row.inscripcion.persona.tipo_documento(),
                                        'carrera': row.inscripcion.carrera.__str__(),
                                        'promedio': row.promedio,
                                        'estado_gratuidad': row.inscripcion.estado_gratuidad,
                                        'estado_gratuidad_texto': row.inscripcion.get_estado_gratuidad_display(),
                                        'genero': row.inscripcion.persona.sexo.__str__() if row.inscripcion.persona.sexo is not None else 'S/N',
                                        'telefonos': ','.join(row.inscripcion.persona.lista_telefonos()),
                                        'direccion': row.inscripcion.persona.direccion_completa(),
                                        'emails': row.inscripcion.persona.lista_emails(),
                                        'foto':foto_url,
                                        'prioridad':row.prioridad
                                        },
                                       {
                                           'telefonos': ','.join(row.inscripcion.persona.lista_telefonos()),
                                           'direccion': row.inscripcion.persona.direccion_completa(),
                                           'emails': row.inscripcion.persona.lista_emails(),
                                       },
                                       {
                                        'estado_gratuidad': row.inscripcion.estado_gratuidad,
                                        'estado_gratuidad_texto': row.inscripcion.get_estado_gratuidad_display(),
                                        'matriculaactual': jsonmatriculaactual,
                                        'matriculaanterior': jsonmatriculaanterior,
                                        'becatipo_id': row.becatipo_id,
                                        'session': row.inscripcion.sesion.nombre,
                                       },
                                       row.inscripcion.mi_nivel().__str__(),
                                       {
                                           "id": row.id,
                                           "prioridad": row.prioridad,
                                           "nombre": row.inscripcion.persona.__str__()
                                       },
                                       ])
                    return JsonResponse({"result": "ok", "data": aaData, "iTotalRecords": tCount, "iTotalDisplayRecords": tCount})
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al cargar los datos. %s" % ex.__str__(), "data": [], "iTotalRecords": 0, "iTotalDisplayRecords": 0})

        elif action == 'changeEstadoPrioridad':
            try:
                id = request.POST.get('id', None)
                prioridad = request.POST.get('prioridad', '0') == '1'
                try:
                    ePreInscripcionBeca = PreInscripcionBeca.objects.get(pk=id)
                except ObjectDoesNotExist:
                    raise NameError('No se encontro pre-inscripción')
                ePreInscripcionBeca.prioridad = prioridad
                ePreInscripcionBeca.save(request)
                if prioridad:
                    log(u'Da prioridad a pre-inscripción: %s' % persona, request, "add")
                else:
                    log(u'Quita prioridad a pre-inscripción: %s' % persona, request, "add")
                return JsonResponse({"result": True, "mensaje": u"Se cambio estado de prioridad correctamente"})
            except Exception as ex:
                msg = str(ex)
                transaction.set_rollback(True)
                return JsonResponse({"result": False, "mensaje": u"%s" % msg})

        elif action == 'enviarcorreopreseleccion':
            try:
                noti = Notificacion(cuerpo=f'Enviando correo de notificacion a becados del periodo: {periodo}',
                                    titulo=f'Envio de notificación por correo de becados en periodo {periodo}, en proceso', destinatario=persona,
                                    url='/adm_becas?action=listasolicitudes',
                                    prioridad=1, app_label='SGA',
                                    fecha_hora_visible=datetime.now() + timedelta(days=1), tipo=2,
                                    en_proceso=True)
                noti.save(request)
                GenerateBackground(request=request, data=data, periodoactual=periodo, periodoanterior=anterior, noti=noti).start()
                return JsonResponse({"result": 'ok', "mensaje": u"Se procede a enviar correos de notificación a los becados"})
            except Exception as ex:
                msg = str(ex)
                transaction.set_rollback(True)
                return JsonResponse({"result": 'bad', "mensaje": u"%s" % msg})

        elif action == 'adicionarprebecado':
            try:
                form = AdicionarPreBecadoForm(request.POST)
                if form.is_valid():
                    becaperiodo = periodo.becaperiodo_set.filter(status=True).first()
                    persona = form.cleaned_data['persona']
                    becatipo = form.cleaned_data['becatipo']
                    observacion = form.cleaned_data['observacion']
                    inscripcion = Inscripcion.objects.filter(persona=persona, status=True, activo=True,
                                                 matricula__nivel__periodo=periodo,
                                                 matricula__status=True,
                                                 matricula__retiradomatricula=False,
                                                 coordinacion__id__in=[1, 2, 3, 4, 5] ).first()

                    matricula_periodo_anterior = Matricula.objects.filter(inscripcion__persona=persona, nivel__periodo=anterior, status=True).first()
                    matricula = Matricula.objects.filter(inscripcion__persona=persona, nivel__periodo=periodo, status=True).first()
                    if inscripcion is None:
                        raise NameError('No se encontro inscripción activa para el periodo actual')

                    if PreInscripcionBeca.objects.filter(inscripcion=inscripcion, periodo=periodo).exists():
                        raise NameError('Ya existe una pre-inscripción para el periodo actual')

                    orden = PreInscripcionBeca.objects.filter(periodo=periodo, becatipo=becatipo).aggregate(Max('orden'))

                    preinscripcionbeca = PreInscripcionBeca(
                        inscripcion=inscripcion,
                        matricula=matricula,
                        promedio=matricula_periodo_anterior.promedionotas,
                        becatipo=becatipo,
                        fecha=datetime.now().date(),
                        periodo=periodo,
                        cumplerequisitos=True,
                        orden=orden['orden__max'] + 1 if orden['orden__max'] else 1,
                        observacion=observacion,
                        becaperiodo=becaperiodo)
                    preinscripcionbeca.save(request)
                    log(u'Adiciona pre-inscripción beca: %s' % persona, request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError("Error al guardar los datos")
            except Exception as ex:
                msg = str(ex)
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error %s" % msg})

        elif action == 'aprobarsolicitudesmasivo':
            try:
                url_path = 'http://127.0.0.1:8000'
                if not DEBUG:
                    url_path = 'https://sga.unemi.edu.ec'
                becassolicitudes = BecaSolicitud.objects.filter(periodocalifica__isnull=False, status=True,periodo=periodo,
                                                                  becaasignacion__isnull=True,
                                                                  inscripcion__persona__personadocumentopersonal__isnull=False,
                                                                  inscripcion__persona__personadocumentopersonal__status=True,
                                                                  inscripcion__persona__cuentabancariapersona__isnull=False,
                                                                  inscripcion__persona__cuentabancariapersona__activapago=True,
                                                                  inscripcion__persona__cuentabancariapersona__status=True).exclude(
                                                                    Q(inscripcion__persona__personadocumentopersonal__estadocedula__in=[3,4,6])|
                                                                    Q(inscripcion__persona__cuentabancariapersona__estadorevision__in=[3,4,6])).order_by('id').distinct()
                cont=0
                for becasolicitud in becassolicitudes:
                    aData = {}
                    aData['becasolicitud']=becasolicitud
                    aData['matricula'] = matricula = becasolicitud.obtener_matricula()
                    ePersona = becasolicitud.inscripcion.persona
                    ePeriodo = becasolicitud.periodo
                    eInscripcion = becasolicitud.inscripcion
                    username = ePersona.usuario.username
                    aData['configuracionbecatipoperiodo'] = becatipoconfiguracion = becasolicitud.obtener_configuracionbecatipoperiodo()
                    becaasignacion = BecaAsignacion.objects.filter(solicitud=becasolicitud, status=True).first()
                    filename = f'acta_compromiso_{eInscripcion.id}_{ePeriodo.id}_{becasolicitud.id}'
                    filenametemp = os.path.join(os.path.join(SITE_STORAGE, 'media', 'becas', 'temp', 'actas_compromisos', username, filename + '.pdf'))
                    filenameqrtemp = os.path.join(os.path.join(SITE_STORAGE, 'media', 'becas', 'temp', 'actas_compromisos', username, 'qrcode', filename + '.png'))
                    existebeca = BecaAsignacion.objects.filter(solicitud__periodo=ePeriodo,solicitud__inscripcion=eInscripcion, status=True).exists()
                    if not becaasignacion and becatipoconfiguracion and not existebeca:
                        if becasolicitud.becatipo_id == 23:
                            beneficios = becasolicitud.becasolicitudnecesidad_set.all().first()
                            beneficio = beneficios.necesidad

                        folder = os.path.join(os.path.join(SITE_STORAGE, 'media', 'becas', 'actas_compromisos', username, ''))
                        folder2 = os.path.join(os.path.join(SITE_STORAGE, 'media', 'becas', 'actas_compromisos', username, 'qrcode', ''))
                        aData['aceptobeca'] = True
                        aData['url_qr'] = rutaimg = folder2 + filename + '.png'
                        aData['rutapdf'] = rutapdf = folder + filename + '.pdf'
                        aData['url_pdf'] = url_pdf = f'{url_path}/media/becas/actas_compromisos/{username}/{filename}.pdf'
                        if os.path.isfile(rutapdf):
                            os.remove(rutapdf)
                        elif os.path.isfile(rutaimg):
                            os.remove(rutaimg)
                        os.makedirs(folder, exist_ok=True)
                        os.makedirs(folder2, exist_ok=True)
                        firma = f'ACEPTADO POR: {eInscripcion.persona.__str__()}\nUSUARIO:{username}\nFECHA: {datetime.utcnow()}\nACEPTO EN: sga.unemi.edu.ec\nDOCUMENTO:{url_pdf}'
                        url = pyqrcode.create(firma)
                        imageqr = url.png(rutaimg, 16, '#000000')
                        aData['version'] = datetime.now().strftime('%Y%m%d_%H%M%S')
                        aData['image_qrcode'] = f'{url_path}/media/becas/actas_compromisos/{username}/qrcode/{filename}.png'
                        aData['fechaactual'] = datetime.now()
                        url_acta = f'becas/actas_compromisos/{username}/{filename}.pdf'
                        rutapdf = folder + filename + '.pdf'

                        valida, pdf, result = conviert_html_to_pdfsaveqr_generico(request, 'alu_becas/actav2.html', {
                            'pagesize': 'A4',
                            'data': aData,

                        }, folder, filename + '.pdf')
                        if valida:
                            becasolicitud.archivoactacompromiso = url_acta
                            becasolicitud.becaaceptada = 2
                            becasolicitud.becaasignada = 2
                            becasolicitud.save(request)
                            # AYUDAS ECONOMICAS COVID
                            if becasolicitud.becatipo.id == 23:
                                modalidad = becasolicitud.inscripcion.carrera.modalidad
                                montobeca = 270.36 if modalidad != 3 else 44.80
                                beca = BecaAsignacion(solicitud=becasolicitud,
                                                      montomensual=montobeca,
                                                      cantidadmeses=1,
                                                      montobeneficio=montobeca,
                                                      fecha=datetime.now().date(),
                                                      activo=True,
                                                      grupopago=None,
                                                      tipo=1,
                                                      notificar=True,
                                                      estadobeca=None,
                                                      infoactualizada=False,
                                                      cargadocumento=True)
                                beca.save(request)
                                detalleuso = BecaDetalleUtilizacion(asignacion=beca,
                                                                    utilizacion_id=8 if beneficio == 1 else 9,
                                                                    personaaprueba=ePersona,
                                                                    archivo=None,
                                                                    fechaaprueba=datetime.now(),
                                                                    fechaarchivo=None,
                                                                    estado=2,
                                                                    observacion='GENERADO AUTOMÁTICAMENTE')
                                detalleuso.save(request)
                                persona = becasolicitud.inscripcion.persona
                                if persona.documentos_personales():
                                    # SI TIENEN DOCUMENTACION VALIDADA
                                    if persona.cedula_solicitante_representante_validadas():
                                        beca.infoactualizada = True
                                        beca.cargadocumento = False
                                        beca.save(request)
                            else:
                                # OTROS TIPOS DE BECA
                                # if solicitud.becatipo_id == 16 or solicitud.becatipo_id == 17:
                                #     montobeca = 110.04
                                # else:
                                #     montobeca = 184.08
                                montobeca = becatipoconfiguracion.becamonto
                                meses = becatipoconfiguracion.becameses
                                montomensual = becatipoconfiguracion.monto_x_mes()
                                beca = BecaAsignacion(solicitud=becasolicitud,
                                                      montomensual=montomensual,
                                                      cantidadmeses=meses,
                                                      montobeneficio=montobeca,
                                                      fecha=datetime.now().date(),
                                                      activo=True,
                                                      grupopago=None,
                                                      tipo=becasolicitud.tiposolicitud,
                                                      notificar=True,
                                                      estadobeca=None,
                                                      infoactualizada=False,
                                                      cargadocumento=True)
                                beca.save(request)
                                # VERIFICAR SI TIENE DOCUMENTACION APROBADA DE UN PROCESO ANTERIOR
                                # persona = becasolicitud.inscripcion.persona
                                # if persona.documentos_personales():
                                #     # SI TIENEN DOCUMENTACION VALIDADA
                                #     if persona.cedula_solicitante_representante_validadas():
                                #         beca.infoactualizada = True
                                #         beca.cargadocumento = False
                                #         beca.save(request)

                                if ePersona.cuentabancaria():
                                    beca.infoactualizada = True
                                    beca.cargadocumento = False
                                    beca.save(request)
                            # url_actafirmada = becasolicitud.archivoactacompromiso.url
                            cont+=1
                            log(u'Aceptó la beca estudiantil: %s' % (ePersona), request, "edit")

                mensaje = f"Se aceptaron correctamente las becas estudiantiles de {cont}"
                return JsonResponse({"result": 'ok', 'showSwal': True, "mensaje": mensaje, 'titulo': 'Solicitudes de becas aceptadas'})
            except Exception as ex:
                msg = str(ex)
                transaction.set_rollback(True)
                return JsonResponse({"result": 'bad', "mensaje": u"%s" % msg})

        elif action == 'rechazarsolicitudesmasivo':
            try:
                import time
                becassolicitudes = BecaSolicitud.objects.filter((Q(inscripcion__persona__personadocumentopersonal__isnull=True) |
                                                                 Q(inscripcion__persona__cuentabancariapersona__isnull=True) |
                                                                 Q(inscripcion__persona__personadocumentopersonal__estadocedula__in=[3,6]) |
                                                                 Q(inscripcion__persona__cuentabancariapersona__estadorevision__in=[3,6])),
                                                                periodocalifica__isnull=False, status=True, periodo=periodo, becaaceptada=1,
                                                                becaasignacion__isnull=True).order_by('id').distinct()
                for becasolicitud in becassolicitudes:
                    if not becasolicitud.cumple_todos_documentos_requeridos_aprobados():
                        ePersona = becasolicitud.inscripcion.persona
                        becatipoconfiguracion = becasolicitud.obtener_configuracionbecatipoperiodo()
                        becasolicitud.becaaceptada = 3
                        becasolicitud.observacion = 'Rechazó la beca estudiantil'
                        becasolicitud.save(request)
                        # AL RECHAZAR EL ESTUDIANTE SU BECA SE SELEECIONARA DE LA PREDATA LOS SIGUIENTES BECARIOS QUE NO FUERON SELECCIONADOS
                        cantidad_estudiantes_becados = BecaSolicitud.objects.values('id').filter(status=True, periodo_id=becasolicitud.periodo_id).exclude(becaaceptada=3).count()
                        cantidad_limite_becados = becatipoconfiguracion.becaperiodo.limitebecados
                        preinscripcion_rechazo = PreInscripcionBeca.objects.filter(inscripcion=becasolicitud.inscripcion, status=True, periodo=becasolicitud.periodo).first()

                        becados_rechazados = BecaSolicitud.objects.values('inscripcion_id').filter(periodo_id=becasolicitud.periodo_id, becaaceptada=3)
                        while cantidad_estudiantes_becados < cantidad_limite_becados:
                            preinscripcionbeca = PreInscripcionBeca.objects.filter(seleccionado=False, status=True, periodo_id=becasolicitud.periodo_id).exclude(
                                inscripcion_id__in=becados_rechazados).order_by('orden').first()
                            if (preinscripcionbeca_p := PreInscripcionBeca.objects.filter(seleccionado=False, status=True, prioridad=True, periodo_id=becasolicitud.periodo_id).exclude(
                                    inscripcion_id__in=becados_rechazados).order_by('orden').first()) is not None:
                                preinscripcionbeca = preinscripcionbeca_p
                            if preinscripcionbeca and not BecaSolicitud.objects.filter(status=True, inscripcion=preinscripcionbeca.inscripcion, periodo=becasolicitud.periodo).exists():
                                becado = BecaSolicitud(inscripcion=preinscripcionbeca.inscripcion,
                                                       becatipo=preinscripcionbeca.becatipo,
                                                       periodo=becasolicitud.periodo,
                                                       periodocalifica=becasolicitud.periodocalifica,
                                                       estado=1,
                                                       tiposolicitud=preinscripcionbeca.tipo_renovacion_nueva(becasolicitud.periodocalifica),
                                                       observacion=f'SEMESTRE REGULAR {preinscripcionbeca.periodo.nombre}')
                                becado.save(usuario_id=1)
                                recorrido = BecaSolicitudRecorrido.objects.filter(solicitud=becado, estado=1).first()
                                preinscripcionbeca.seleccionado = True
                                preinscripcionbeca.save(usuario_id=1)
                                if recorrido is None:
                                    recorrido = BecaSolicitudRecorrido(solicitud=becado,
                                                                       observacion="SOLICITUD AUTOMÁTICA",
                                                                       estado=1)
                                    recorrido.save(usuario_id=1)
                                    recorrido = BecaSolicitudRecorrido.objects.filter(solicitud=becado, estado=4).first()
                                    # REGISTRO EN ESTADO DE REVISION
                                    if recorrido is None:
                                        becado.estado = 4
                                        becado.save(usuario_id=1)
                                        recorrido = BecaSolicitudRecorrido(solicitud=becado,
                                                                           observacion="EN REVISION",
                                                                           estado=4)
                                        recorrido.save(usuario_id=1)
                                        # log(u'Agrego recorrido de beca: %s' % recorrido, request, "add")

                                    recorrido = BecaSolicitudRecorrido.objects.filter(solicitud=becado, estado=2).first()
                                    if recorrido is None:
                                        becado.estado = 2
                                        becado.becaaceptada = 1
                                        becado.save(usuario_id=1)
                                        recorrido = BecaSolicitudRecorrido(solicitud=becado,
                                                                           observacion="PENDIENTE DE ACEPTACIÓN O RECHAZO",
                                                                           estado=2)
                                        recorrido.save(usuario_id=1)
                                        # log(u'Agrego recorrido de beca: %s' % recorrido, request, "add")
                                        notificacion = Notificacion(titulo=f"Solicitud de Beca en Revisión - {preinscripcionbeca.periodo.nombre}",
                                                                    cuerpo=f"Solicitud de aplicación a la beca por {becado.becatipo.nombre.upper()} para el periodo académico {becado.periodo.nombre} ha sido {'APROBADA' if becado.estado == 2 else 'RECHAZADA'}",
                                                                    destinatario=preinscripcionbeca.inscripcion.persona,
                                                                    url="/alu_becas",
                                                                    content_type=None,
                                                                    object_id=None,
                                                                    prioridad=2,
                                                                    perfil=preinscripcionbeca.inscripcion.perfil_usuario(),
                                                                    app_label='SIE',  # request.session['tiposistema'],
                                                                    fecha_hora_visible=datetime.now() + timedelta(days=2)
                                                                    )
                                        notificacion.save(usuario_id=1)
                                        preinscripcion_rechazo.seleccionado = False
                                        preinscripcion_rechazo.save(request)
                                        lista_envio_masivo = []
                                        tituloemail = "Beca estudiantil"
                                        data = {'sistema': u'SGA - UNEMI',
                                                'fase': 'AR',
                                                'tipobeca': preinscripcionbeca.becatipo.nombre.upper(),
                                                'fecha': datetime.now().date(),
                                                'hora': datetime.now().time(),
                                                'saludo': 'Estimada' if becado.inscripcion.persona.sexo_id == 1 else 'Estimado',
                                                'estado': 'APROBADA' if becado.estado == 2 else "RECHAZADA",
                                                'estudiante': becado.inscripcion.persona.nombre_minus(),
                                                'autoridad2': '',
                                                'observaciones': '',
                                                'periodo': becado.periodo.nombre,
                                                't': miinstitucion()
                                                }
                                        list_email = becado.inscripcion.persona.lista_emails_envio()
                                        plantilla = "emails/notificarestadosolicitudbeca.html"
                                        send_html_mail(tituloemail,
                                                       plantilla,
                                                       data,
                                                       list_email,
                                                       [],
                                                       cuenta=CUENTAS_CORREOS[0][1]
                                                       )
                                        time.sleep(3)
                                        cantidad_estudiantes_becados = BecaSolicitud.objects.values('id').filter(status=True, periodo_id=becasolicitud.periodo_id).exclude(becaaceptada=3).count()
                            else:
                                break
                        log(u'Rechazó la beca estudiantil: %s' % ePersona, request, "edit")

                mensaje = f"Se rechazo correctamente {len(becassolicitudes)} solicitude de becas estudiantiles"
                return JsonResponse({"result": 'ok','showSwal':True, "mensaje": mensaje,'titulo':'Solicitudes de becas rechazados'})
            except Exception as ex:
                msg = str(ex)
                transaction.set_rollback(True)
                return JsonResponse({"result": 'bad', "mensaje": u"%s" % msg})

        elif action == 'anularasignacionbecamasivo':
            try:
                import time
                form = BecaSolicitudAnulacionForm(request.POST, request.FILES)
                becas = BecaAsignacion.objects.filter((Q(Q(solicitud__inscripcion__persona__personadocumentopersonal__isnull=True) & Q(solicitud__inscripcion__persona__cuentabancariapersona__isnull=True)) |
                                                      Q(Q(solicitud__inscripcion__persona__personadocumentopersonal__estadocedula__in=[3,6]) & Q(solicitud__inscripcion__persona__cuentabancariapersona__isnull=True)) |
                                                      Q(solicitud__inscripcion__persona__cuentabancariapersona__estadorevision__in=[3,6])),
                                                      status=True, solicitud__periodo=periodo).exclude(estadobeca=3).order_by('id').distinct()
                if not form.is_valid():
                    raise NameError('Llene todos los campos requeridos.')
                for beca in becas:
                    if not beca.solicitud.cumple_todos_documentos_requeridos_aprobados() and not beca.solicitudpagobecadetalle_set.filter(status=True).exists():
                        anular_beca(request, beca, form)

                mensaje = f"Se anulo correctamente {len(becas)} asignaciones de becas estudiantiles"
                messages.success(request, f"{mensaje}")
                return JsonResponse({"result": 'ok', "mensaje": mensaje,})
            except Exception as ex:
                msg = str(ex)
                transaction.set_rollback(True)
                return JsonResponse({"result": 'bad', "mensaje": u"%s" % msg})

        return JsonResponse({"result": "bad", "mensaje": u"Solicitud Incorrecta."})

    else:
        if 'action' in request.GET:
            action = request.GET['action']

            if action == 'mostrardocumentos':
                try:
                    data = {}
                    data['beca'] = beca = BecaAsignacion.objects.get(solicitud_id=int(request.GET['id']))
                    data['docvalidado'] = beca.solicitud.inscripcion.persona.cedula_solicitante_representante_validadas() if beca.solicitud.becatipo.id != 16 else beca.solicitud.inscripcion.persona.cedula_solicitante_representante_validadasbpn()
                    data['documentos'] = beca.solicitud.inscripcion.persona.documentos_personales()
                    data['cuentabancaria'] = persona.cuentabancaria()
                    idperiodobeca = beca.solicitud.periodo.id
                    isPersonaExterior = beca.solicitud.inscripcion.persona.ecuatoriano_vive_exterior()
                    data["isPersonaExterior"] = isPersonaExterior
                    data['inscripcion'] =  beca.solicitud.inscripcion
                    data['documentopersonal'] = beca.solicitud.inscripcion.persona.personadocumentopersonal_set.filter(status=True).first()
                    data['cuentabancaria'] = cuentabancaria = persona.cuentabancaria_becas()
                    data['perfil'] = beca.solicitud.inscripcion.persona.mi_perfil()
                    data['deportista'] = beca.solicitud.inscripcion.persona.deportistapersona_set.filter(status=True).first()
                    data['cuentabancaria'] = beca.solicitud.inscripcion.persona.cuentabancaria()
                    comprobantes = beca.becacomprobanterevision_set.filter(status=True)
                    data['comprobante'] = comprobantes[0] if comprobantes else None

                    perilinscripcion = None
                    if beca.solicitud.inscripcion.persona.perfilinscripcion_set.filter(status=True).exists():
                        perilinscripcion = beca.solicitud.inscripcion.persona.perfilinscripcion_set.filter(
                            status=True).first()
                    data['perilinscripcion'] = perilinscripcion

                    template = get_template("alu_becas/mostrardocumentos.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content})
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

            elif action == 'mostrarcomprobantes':
                try:
                    data = {}
                    fechaactual = datetime.now().date()
                    data['beca'] = beca = BecaAsignacion.objects.get(solicitud_id=int(request.GET['id']))
                    # data['comprobantes'] = comprobantes = beca.becacomprobanteventa_set.filter(status=True).order_by('id')

                    revisioncomprobante = beca.becacomprobanterevision_set.filter(status=True)[
                        0] if beca.becacomprobanterevision_set.filter(status=True).exists() else None
                    data['comprobantes'] = comprobantes = revisioncomprobante.becacomprobanteventa_set.filter(
                        status=True).order_by('id') if revisioncomprobante else None

                    detallepago = beca.solicitudpagobecadetalle_set.all()[0]
                    fechaacredita = datetime.strptime(str(detallepago.fechaacredita)[:10], "%Y-%m-%d").date()
                    diastranscurridos = abs((fechaactual - fechaacredita).days)

                    PLAZO = 15

                    if not comprobantes:
                        data['permitiragregar'] = True if diastranscurridos <= PLAZO else False
                    else:
                        cargado = comprobantes.filter(estadorevisiondbu=1).count()
                        rechazado = comprobantes.filter(estadorevisiondbu__in=[3, 6]).count()
                        data['permitiragregar'] = True if rechazado > 0 or cargado > 0 else False

                    data['mostrarmensajeplazo'] = True if not comprobantes and diastranscurridos > PLAZO else False
                    data['fechaacredita'] = fechaacredita
                    data['fechaactual'] = fechaactual
                    data['dias'] = diastranscurridos + 1

                    data['modolectura'] = True if request.GET['modo'] == 'L' else False

                    template = get_template("alu_becas/mostrarcomprobantes.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content})
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

            elif action == 'matrizsolicitudes':
                try:
                    __author__ = 'Unemi'
                    desde = datetime.combine(convertir_fecha(request.GET['desde']), time.min)
                    hasta = datetime.combine(convertir_fecha(request.GET['hasta']), time.max)
                    tipobeca = int(request.GET['tipobeca'])
                    beca = BecaTipo.objects.get(pk=tipobeca)

                    titulo = easyxf(
                        'font: name Times New Roman, color-index black, bold on , height 350; alignment: horiz centre')
                    titulo2 = easyxf(
                        'font: name Times New Roman, color-index black, bold on , height 250; alignment: horiz centre')
                    fuentecabecera = easyxf(
                        'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
                    fuentenormal = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
                    fuentemoneda = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right',
                        num_format_str=' "$" #,##0.00')
                    fuentefecha = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz center',
                        num_format_str='yyyy-mm-dd')
                    fuentenumerodecimal = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right',
                        num_format_str='#,##0.00')
                    fuentenumeroentero = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right')

                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('Solicitudes')
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=matriz_solicitud_beca_' + random.randint(1,
                                                                                                                     10000).__str__() + '.xls'

                    ws.write_merge(0, 0, 0, 12, 'UNIVERSIDAD ESTATAL DE MILAGRO', titulo)
                    ws.write_merge(1, 1, 0, 12, 'DIRECCIÓN DE BIENESTAR UNIVERSITARIO', titulo2)
                    ws.write_merge(2, 2, 0, 12, 'MATRIZ DE SOLICITUDES DE BECAS', titulo2)
                    ws.write_merge(3, 3, 0, 12,
                                   'DESDE:   ' + str(convertir_fecha(request.GET['desde'])) + '     HASTA:   ' + str(
                                       convertir_fecha(request.GET['hasta'])), titulo2)
                    ws.write_merge(4, 4, 0, 12, 'TIPO DE BECA: ' + beca.nombre.upper(), titulo2)
                    ws.write_merge(5, 5, 0, 12, 'PERIODO: ' + periodo.nombre, titulo2)

                    row_num = 7

                    columns = [
                        (u"# SOLICITUD", 3000),
                        (u"FECHA SOLICITUD", 3000),
                        (u"ESTADO", 3000),
                        (u"NOMBRES COMPLETOS", 10000),
                        (u"IDENTIFICACIÓN", 5000),
                        (u"FECHA NACIMIENTO", 3000),
                        (u"FACULTAD", 7000),
                        (u"CARRERA", 7000),
                        (u"MODALIDAD", 7000),
                        (u"SEMESTRE ANTERIOR", 4000),
                        (u"PROMEDIO ANTERIOR", 3000),
                        (u"ASISTENCIA ANTERIOR", 3000),
                        (u"E-MAIL PERSONAL", 7000),
                        (u"E-MAIL UNEMI", 7000),
                        (u"PROVINCIA", 5000),
                        (u"CANTÓN", 5000),
                        (u"DIRECCIÓN", 20000)
                    ]

                    if tipobeca == 18 or tipobeca == 19:
                        columns.append((u"GRUPO SOCIOECONÓMICO" if tipobeca == 18 else u"TIPO DISCAPACIDAD", 5000))

                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], fuentecabecera)
                        ws.col(col_num).width = columns[col_num][1]

                    # if periodo.id == 110:
                    #     requisitos = BecaRequisitos.objects.filter(Q(becatipo__isnull=True) | Q(becatipo_id=tipobeca) , periodo=periodo, status=True, vigente=True).exclude(numero=7).order_by('numero')
                    # else:
                    #     requisitos = BecaRequisitos.objects.filter(Q(becatipo__isnull=True) | Q(becatipo_id=tipobeca) , periodo=periodo, status=True, vigente=True).order_by('numero')

                    # col_num += 1
                    # nombrereq = ""
                    # for r in requisitos:
                    #     if r.numero == 1:
                    #         nombrereq = r.nombre
                    #         continue
                    #     elif r.numero == 2:
                    #         nombrereq = nombrereq + " / " + r.nombre
                    #     else:
                    #         nombrereq = r.nombre
                    #
                    #     ws.write(row_num, col_num, nombrereq, fuentecabecera)
                    #
                    #     ws.col(col_num).width = 8000
                    #     col_num += 1

                    row_num = 7

                    solicitudes = BecaSolicitud.objects.filter(fecha_creacion__range=(desde, hasta), status=True,
                                                               periodo=periodo, becatipo_id=tipobeca).exclude(
                        estado__in=[5, 8]).order_by('-id')

                    for s in solicitudes:
                        row_num += 1

                        p = Persona.objects.get(pk=s.inscripcion.persona_id)

                        if Matricula.objects.values('id').filter(inscripcion__persona=p,
                                                                 nivel__periodo=s.periodocalifica, status=True,
                                                                 estado_matricula__in=[2, 3]).count() == 1:
                            umat = Matricula.objects.get(inscripcion__persona=p, nivel__periodo=s.periodocalifica,
                                                         status=True, estado_matricula__in=[2, 3])
                        else:
                            matriculas = Matricula.objects.filter(inscripcion__persona=p, status=True,
                                                                  estado_matricula__in=[2, 3],
                                                                  nivel__periodo=s.periodocalifica).order_by(
                                '-nivel__periodo__fin')
                            for x in matriculas:
                                if x.materias_aprobadas_todas():
                                    umat = Matricula.objects.get(pk=x.id, nivel__periodo=s.periodocalifica, status=True,
                                                                 estado_matricula__in=[2, 3])
                                    break

                        ws.write(row_num, 0, str(s.id), fuentenumeroentero)
                        ws.write(row_num, 1, str(s.fecha_creacion)[:10], fuentefecha)
                        ws.write(row_num, 2, s.get_estado_display(), fuentenormal)
                        ws.write(row_num, 3, p.nombre_completo_inverso(), fuentenormal)
                        ws.write(row_num, 4, p.identificacion(), fuentenormal)
                        ws.write(row_num, 5, str(p.nacimiento)[:10], fuentefecha)
                        ws.write(row_num, 6, str(umat.nivel.coordinacion()), fuentenormal)
                        ws.write(row_num, 7, str(umat.inscripcion.carrera), fuentenormal)
                        ws.write(row_num, 8, str(umat.inscripcion.modalidad), fuentenormal)
                        ws.write(row_num, 9, umat.nivelmalla.nombre, fuentenormal)
                        ws.write(row_num, 10, umat.promedio_nota_dbu(), fuentenumerodecimal)
                        ws.write(row_num, 11, umat.promedio_asistencias_dbu(), fuentenumerodecimal)
                        ws.write(row_num, 12, p.email, fuentenormal)
                        ws.write(row_num, 13, p.emailinst, fuentenormal)
                        ws.write(row_num, 14, str(p.provincia), fuentenormal)
                        ws.write(row_num, 15, str(p.canton), fuentenormal)
                        ws.write(row_num, 16, p.direccion_corta(), fuentenormal)
                        if tipobeca == 18 or tipobeca == 19:
                            if tipobeca == 18:
                                ws.write(row_num, 17, p.mi_ficha().grupoeconomico.codigo, fuentenormal)
                            else:
                                ws.write(row_num, 17,
                                         p.mi_perfil().tipodiscapacidad.nombre if p.mi_perfil().tipodiscapacidad else '',
                                         fuentenormal)

                    wb.save(response)
                    return response
                except Exception as ex:
                    print("Error...")
                    pass

            elif action == 'matrizsolicitudesdesistidas':
                try:
                    __author__ = 'Unemi'

                    titulo = easyxf(
                        'font: name Times New Roman, color-index black, bold on , height 350; alignment: horiz centre')
                    titulo2 = easyxf(
                        'font: name Times New Roman, color-index black, bold on , height 250; alignment: horiz centre')
                    fuentecabecera = easyxf(
                        'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
                    fuentenormal = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
                    fuentemoneda = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right',
                        num_format_str=' "$" #,##0.00')
                    fuentefecha = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz center',
                        num_format_str='yyyy-mm-dd')
                    fuentenumerodecimal = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right',
                        num_format_str='#,##0.00')
                    fuentenumeroentero = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right')

                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('SolDesistidas')
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=solicitud_beca_desistida_' + random.randint(
                        1, 10000).__str__() + '.xls'

                    ws.write_merge(0, 0, 0, 12, 'UNIVERSIDAD ESTATAL DE MILAGRO', titulo)
                    ws.write_merge(1, 1, 0, 12, 'DIRECCIÓN DE BIENESTAR UNIVERSITARIO', titulo2)
                    ws.write_merge(2, 2, 0, 12, 'LISTADO DE SOLICITUDES DE BECAS DESISTIDAS', titulo2)
                    # ws.write_merge(3, 3, 0, 12, 'DESDE:   ' + str(convertir_fecha(request.GET['desde'])) + '     HASTA:   ' + str(convertir_fecha(request.GET['hasta'])), titulo2)
                    # ws.write_merge(4, 4, 0, 12, 'TIPO DE BECA: ' + beca.nombre.upper(), titulo2)

                    row_num = 4
                    columns = [
                        (u"# SOLICITUD", 3000),
                        (u"FECHA DESISTIÓ", 3000),
                        (u"ESTADO", 3000),
                        (u"NOMBRES COMPLETOS", 10000),
                        (u"IDENTIFICACIÓN", 5000),
                        (u"FECHA NACIMIENTO", 3000),
                        (u"FACULTAD", 7000),
                        (u"CARRERA", 7000),
                        (u"MODALIDAD", 7000),
                        (u"NIVEL", 5000),
                        (u"E-MAIL PERSONAL", 7000),
                        (u"E-MAIL UNEMI", 7000),
                        (u"PROVINCIA", 5000),
                        (u"CANTÓN", 5000),
                        (u"DIRECCIÓN", 20000)
                    ]
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], fuentecabecera)
                        ws.col(col_num).width = columns[col_num][1]

                    row_num = 4

                    solicitudes = BecaSolicitud.objects.filter(status=True, periodo=periodo, estado=8).order_by(
                        '-fecha_modificacion')

                    for s in solicitudes:
                        row_num += 1

                        p = Persona.objects.get(pk=s.inscripcion.persona_id)
                        umat = s.inscripcion.matricula_periodo_actual(periodo)[0]

                        ws.write(row_num, 0, str(s.id), fuentenumeroentero)
                        ws.write(row_num, 1, str(s.fecha_modificacion)[:10], fuentefecha)
                        ws.write(row_num, 2, s.get_estado_display(), fuentenormal)
                        ws.write(row_num, 3, p.nombre_completo_inverso(), fuentenormal)
                        ws.write(row_num, 4, p.identificacion(), fuentenormal)
                        ws.write(row_num, 5, str(p.nacimiento)[:10], fuentefecha)
                        ws.write(row_num, 6, str(umat.nivel.coordinacion()), fuentenormal)
                        ws.write(row_num, 7, str(umat.inscripcion.carrera), fuentenormal)
                        ws.write(row_num, 8, str(umat.inscripcion.modalidad), fuentenormal)
                        ws.write(row_num, 9, umat.nivelmalla.nombre, fuentenormal)
                        ws.write(row_num, 10, p.email, fuentenormal)
                        ws.write(row_num, 11, p.emailinst, fuentenormal)
                        ws.write(row_num, 12, str(p.provincia), fuentenormal)
                        ws.write(row_num, 13, str(p.canton), fuentenormal)
                        ws.write(row_num, 14, p.direccion_corta(), fuentenormal)

                    wb.save(response)
                    return response
                except Exception as ex:
                    print("Error...")
                    pass

            elif action == 'matrizsolicitudespendienteaceptar':
                try:
                    __author__ = 'Unemi'

                    titulo = easyxf(
                        'font: name Times New Roman, color-index black, bold on , height 350; alignment: horiz centre')
                    titulo2 = easyxf(
                        'font: name Times New Roman, color-index black, bold on , height 250; alignment: horiz centre')
                    fuentecabecera = easyxf(
                        'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
                    fuentenormal = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
                    fuentemoneda = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right',
                        num_format_str=' "$" #,##0.00')
                    fuentefecha = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz center',
                        num_format_str='yyyy-mm-dd')
                    fuentenumerodecimal = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right',
                        num_format_str='#,##0.00')
                    fuentenumeroentero = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right')

                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('SolDesistidas')
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=beca_pendiente_aceptar_' + random.randint(1,
                                                                                                                      10000).__str__() + '.xls'

                    ws.write_merge(0, 0, 0, 14, 'UNIVERSIDAD ESTATAL DE MILAGRO', titulo)
                    ws.write_merge(1, 1, 0, 14, 'DIRECCIÓN DE BIENESTAR UNIVERSITARIO', titulo2)
                    ws.write_merge(2, 2, 0, 14, 'LISTADO DE BECAS PENDIENTES DE ACEPTAR POR LOS ALUMNOS', titulo2)
                    # ws.write_merge(3, 3, 0, 12, 'DESDE:   ' + str(convertir_fecha(request.GET['desde'])) + '     HASTA:   ' + str(convertir_fecha(request.GET['hasta'])), titulo2)
                    # ws.write_merge(4, 4, 0, 12, 'TIPO DE BECA: ' + beca.nombre.upper(), titulo2)

                    row_num = 4
                    columns = [
                        (u"# SOLICITUD", 3000),
                        (u"FECHA BECA", 3000),
                        (u"ESTADO", 3000),
                        (u"NOMBRES COMPLETOS", 10000),
                        (u"IDENTIFICACIÓN", 5000),
                        (u"FECHA NACIMIENTO", 3000),
                        (u"FACULTAD", 7000),
                        (u"CARRERA", 7000),
                        (u"MODALIDAD", 7000),
                        (u"NIVEL", 5000),
                        (u"E-MAIL PERSONAL", 7000),
                        (u"E-MAIL UNEMI", 7000),
                        (u"TELÉFONO", 7000),
                        (u"CELULAR", 7000),
                        (u"PROVINCIA", 5000),
                        (u"CANTÓN", 5000),
                        (u"DIRECCIÓN", 20000)
                    ]
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], fuentecabecera)
                        ws.col(col_num).width = columns[col_num][1]

                    row_num = 4

                    solicitudes = BecaSolicitud.objects.filter(status=True, periodo=periodo, estado=2,
                                                               becaaceptada=1).order_by('-fecha_modificacion')

                    for s in solicitudes:
                        row_num += 1

                        p = Persona.objects.get(pk=s.inscripcion.persona_id)
                        umat = s.inscripcion.matricula_periodo_actual(periodo)[0]

                        ws.write(row_num, 0, str(s.id), fuentenumeroentero)
                        ws.write(row_num, 1, str(s.fecha_modificacion)[:10], fuentefecha)
                        ws.write(row_num, 2, s.get_estado_display(), fuentenormal)
                        ws.write(row_num, 3, p.nombre_completo_inverso(), fuentenormal)
                        ws.write(row_num, 4, p.identificacion(), fuentenormal)
                        ws.write(row_num, 5, str(p.nacimiento)[:10], fuentefecha)
                        ws.write(row_num, 6, str(umat.nivel.coordinacion()), fuentenormal)
                        ws.write(row_num, 7, str(umat.inscripcion.carrera), fuentenormal)
                        ws.write(row_num, 8, str(umat.inscripcion.modalidad), fuentenormal)
                        ws.write(row_num, 9, umat.nivelmalla.nombre, fuentenormal)
                        ws.write(row_num, 10, p.email, fuentenormal)
                        ws.write(row_num, 11, p.emailinst, fuentenormal)
                        ws.write(row_num, 12, p.telefono_conv, fuentenormal)
                        ws.write(row_num, 13, p.telefono, fuentenormal)
                        ws.write(row_num, 14, str(p.provincia), fuentenormal)
                        ws.write(row_num, 15, str(p.canton), fuentenormal)
                        ws.write(row_num, 16, p.direccion_corta(), fuentenormal)

                    wb.save(response)
                    return response
                except Exception as ex:
                    print("Error...")
                    pass

            elif action == 'listadobeneficiarios':
                try:
                    __author__ = 'Unemi'

                    titulo = easyxf(
                        'font: name Times New Roman, color-index black, bold on , height 350; alignment: horiz centre')
                    titulo2 = easyxf(
                        'font: name Times New Roman, color-index black, bold on , height 250; alignment: horiz centre')
                    fuentecabecera = easyxf(
                        'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
                    fuentenormal = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
                    fuentemoneda = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right',
                        num_format_str=' "$" #,##0.00')
                    fuentefecha = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz center',
                        num_format_str='yyyy-mm-dd')
                    fuentenumerodecimal = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right',
                        num_format_str='#,##0.00')
                    fuentenumeroentero = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right')

                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('Listado Beneficiarios')
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=beca_beneficiarios_' + random.randint(1,
                                                                                                                  10000).__str__() + '.xls'

                    ws.write_merge(0, 0, 0, 19, 'UNIVERSIDAD ESTATAL DE MILAGRO', titulo)
                    ws.write_merge(1, 1, 0, 19, 'DIRECCIÓN DE BIENESTAR UNIVERSITARIO', titulo2)
                    ws.write_merge(2, 2, 0, 19, 'LISTADO DE BENEFICIARIOS DE BECAS', titulo2)
                    # ws.write_merge(3, 3, 0, 12, 'DESDE:   ' + str(convertir_fecha(request.GET['desde'])) + '     HASTA:   ' + str(convertir_fecha(request.GET['hasta'])), titulo2)
                    # ws.write_merge(4, 4, 0, 12, 'TIPO DE BECA: ' + beca.nombre.upper(), titulo2)

                    row_num = 4
                    columns = [
                        (u"# SOLICITUD", 3000),
                        (u"FECHA BECA", 3000),
                        (u"NOMBRES", 10000),
                        (u"APELLIDOS", 10000),
                        (u"IDENTIFICACIÓN", 5000),
                        (u"FECHA NACIMIENTO", 3000),
                        (u"FACULTAD", 7000),
                        (u"CARRERA", 7000),
                        (u"MODALIDAD", 7000),
                        (u"NIVEL", 5000),
                        (u"E-MAIL PERSONAL", 7000),
                        (u"E-MAIL UNEMI", 7000),
                        (u"TELÉFONO", 7000),
                        (u"CELULAR", 7000),
                        (u"PROVINCIA", 5000),
                        (u"CANTÓN", 5000),
                        (u"DIRECCIÓN", 20000),
                        (u"TIPO BECA", 20000),
                        (u"ESTADO BECA", 3000),
                        (u"ESTADO", 3000),
                    ]
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], fuentecabecera)
                        ws.col(col_num).width = columns[col_num][1]

                    row_num = 4

                    solicitudes = BecaSolicitud.objects.filter(status=True, periodo=periodo).order_by(
                        '-fecha_modificacion')

                    for s in solicitudes:
                        row_num += 1

                        p = Persona.objects.get(pk=s.inscripcion.persona_id)
                        umat = s.inscripcion.matricula_periodo_actual(periodo)[0]

                        ws.write(row_num, 0, str(s.id), fuentenumeroentero)
                        ws.write(row_num, 1, str(s.fecha_modificacion)[:10], fuentefecha)
                        ws.write(row_num, 2, p.nombres, fuentenormal)
                        ws.write(row_num, 3, p.apellidos(), fuentenormal)
                        ws.write(row_num, 4, p.identificacion(), fuentenormal)
                        ws.write(row_num, 5, str(p.nacimiento)[:10], fuentefecha)
                        ws.write(row_num, 6, str(umat.nivel.coordinacion()), fuentenormal)
                        ws.write(row_num, 7, str(umat.inscripcion.carrera), fuentenormal)
                        ws.write(row_num, 8, str(umat.inscripcion.modalidad), fuentenormal)
                        ws.write(row_num, 9, umat.nivelmalla.nombre, fuentenormal)
                        ws.write(row_num, 10, p.email, fuentenormal)
                        ws.write(row_num, 11, p.emailinst, fuentenormal)
                        ws.write(row_num, 12, p.telefono_conv, fuentenormal)
                        ws.write(row_num, 13, p.telefono, fuentenormal)
                        ws.write(row_num, 14, str(p.provincia), fuentenormal)
                        ws.write(row_num, 15, str(p.canton), fuentenormal)
                        ws.write(row_num, 16, p.direccion_corta(), fuentenormal)
                        ws.write(row_num, 17, str(s.becatipo), fuentenormal)
                        ws.write(row_num, 18, s.get_becaaceptada_display(), fuentenormal)
                        ws.write(row_num, 19, s.get_estado_display(), fuentenormal)
                    wb.save(response)
                    return response
                except Exception as ex:
                    print("Error...")
                    pass

            elif action == 'matrizsolicitudespendientesrevisar':
                try:
                    __author__ = 'Unemi'

                    titulo = easyxf(
                        'font: name Times New Roman, color-index black, bold on , height 350; alignment: horiz centre')
                    titulo2 = easyxf(
                        'font: name Times New Roman, color-index black, bold on , height 250; alignment: horiz centre')
                    fuentecabecera = easyxf(
                        'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
                    fuentenormal = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
                    fuentemoneda = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right',
                        num_format_str=' "$" #,##0.00')
                    fuentefecha = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz center',
                        num_format_str='yyyy-mm-dd')
                    fuentenumerodecimal = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right',
                        num_format_str='#,##0.00')
                    fuentenumeroentero = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right')

                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('SolDesistidas')
                    response = HttpResponse(content_type="application/ms-excel")
                    response[
                        'Content-Disposition'] = 'attachment; filename=solicitud_beca_pendiente_revisar_' + random.randint(
                        1, 10000).__str__() + '.xls'

                    ws.write_merge(0, 0, 0, 12, 'UNIVERSIDAD ESTATAL DE MILAGRO', titulo)
                    ws.write_merge(1, 1, 0, 12, 'DIRECCIÓN DE BIENESTAR UNIVERSITARIO', titulo2)
                    ws.write_merge(2, 2, 0, 12, 'LISTADO DE SOLICITUDES DE BECAS PENDIENTES DE APROBAR/RECHAZAR',
                                   titulo2)
                    # ws.write_merge(3, 3, 0, 12, 'DESDE:   ' + str(convertir_fecha(request.GET['desde'])) + '     HASTA:   ' + str(convertir_fecha(request.GET['hasta'])), titulo2)
                    # ws.write_merge(4, 4, 0, 12, 'TIPO DE BECA: ' + beca.nombre.upper(), titulo2)

                    row_num = 4
                    columns = [
                        (u"# SOLICITUD", 3000),
                        (u"FECHA SOLICITUD", 3000),
                        (u"ESTADO", 3000),
                        (u"NOMBRES COMPLETOS", 10000),
                        (u"IDENTIFICACIÓN", 5000),
                        (u"FECHA NACIMIENTO", 3000),
                        (u"FACULTAD", 7000),
                        (u"CARRERA", 7000),
                        (u"MODALIDAD", 7000),
                        (u"NIVEL", 5000),
                        (u"E-MAIL PERSONAL", 7000),
                        (u"E-MAIL UNEMI", 7000),
                        (u"PROVINCIA", 5000),
                        (u"CANTÓN", 5000),
                        (u"DIRECCIÓN", 20000)
                    ]
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], fuentecabecera)
                        ws.col(col_num).width = columns[col_num][1]

                    row_num = 4

                    solicitudes = BecaSolicitud.objects.filter(status=True, periodo=periodo, estado__in=[1, 4],
                                                               periodocalifica__isnull=False).order_by(
                        '-fecha_creacion')

                    for s in solicitudes:
                        row_num += 1

                        p = Persona.objects.get(pk=s.inscripcion.persona_id)
                        umat = s.inscripcion.matricula_periodo_actual(periodo)[0]

                        ws.write(row_num, 0, str(s.id), fuentenumeroentero)
                        ws.write(row_num, 1, str(s.fecha_creacion)[:10], fuentefecha)
                        ws.write(row_num, 2, s.get_estado_display(), fuentenormal)
                        ws.write(row_num, 3, p.nombre_completo_inverso(), fuentenormal)
                        ws.write(row_num, 4, p.identificacion(), fuentenormal)
                        ws.write(row_num, 5, str(p.nacimiento)[:10], fuentefecha)
                        ws.write(row_num, 6, str(umat.nivel.coordinacion()), fuentenormal)
                        ws.write(row_num, 7, str(umat.inscripcion.carrera), fuentenormal)
                        ws.write(row_num, 8, str(umat.inscripcion.modalidad), fuentenormal)
                        ws.write(row_num, 9, umat.nivelmalla.nombre, fuentenormal)
                        ws.write(row_num, 10, p.email, fuentenormal)
                        ws.write(row_num, 11, p.emailinst, fuentenormal)
                        ws.write(row_num, 12, str(p.provincia), fuentenormal)
                        ws.write(row_num, 13, str(p.canton), fuentenormal)
                        ws.write(row_num, 14, p.direccion_corta(), fuentenormal)

                    wb.save(response)
                    return response
                except Exception as ex:
                    print("Error...")
                    pass

            elif action == 'becadosdocumentacionvalidada':
                try:
                    __author__ = 'Unemi'
                    desde = datetime.combine(convertir_fecha(request.GET['desde']), time.min)
                    hasta = datetime.combine(convertir_fecha(request.GET['hasta']), time.max)
                    # tipobeca = int(request.GET['tipobeca'])

                    titulo = easyxf(
                        'font: name Times New Roman, color-index black, bold on , height 350; alignment: horiz centre')
                    titulo2 = easyxf(
                        'font: name Times New Roman, color-index black, bold on , height 250; alignment: horiz centre')
                    fuentecabecera = easyxf(
                        'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
                    fuentenormal = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
                    fuentemoneda = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right',
                        num_format_str=' "$" #,##0.00')
                    fuentefecha = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz center',
                        num_format_str='yyyy-mm-dd')
                    fuentenumerodecimal = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right',
                        num_format_str='#,##0.00')
                    fuentenumeroentero = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right')

                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('Becados')
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=becados_docuvalidada_' + random.randint(1,
                                                                                                                    10000).__str__() + '.xls'

                    ws.write_merge(0, 0, 0, 14, 'UNIVERSIDAD ESTATAL DE MILAGRO', titulo)
                    ws.write_merge(1, 1, 0, 14, 'DIRECCIÓN DE BIENESTAR UNIVERSITARIO', titulo2)
                    ws.write_merge(2, 2, 0, 14, 'LISTADO DE BECADOS - DOCUMENTACIÓN VALIDADA', titulo2)
                    ws.write_merge(3, 3, 0, 14,
                                   'DESDE:   ' + str(convertir_fecha(request.GET['desde'])) + '     HASTA:   ' + str(
                                       convertir_fecha(request.GET['hasta'])), titulo2)

                    row_num = 6
                    columns = [
                        (u"# SOLICITUD", 3000),
                        (u"FECHA BECA", 3000),
                        (u"MONTO BECA", 3000),
                        (u"NOMBRES COMPLETOS", 10000),
                        (u"IDENTIFICACIÓN", 5000),
                        (u"FACULTAD", 5000),
                        (u"CARRERA", 5000),
                        (u"MODALIDAD", 5000),
                        (u"NIVEL", 5000),
                        (u"PROVINCIA", 5000),
                        (u"CANTÓN", 5000),
                        (u"PARROQUIA", 5000),
                        (u"DIRECCIÓN", 15000),
                        (u"REFERENCIA", 15000),
                        (u"SECTOR", 15000),
                        (u"# CASA", 5000),
                        (u"E-MAIL PERSONAL", 7000),
                        (u"E-MAIL UNEMI", 7000),
                        (u"TELÉFONO", 5000),
                        (u"CELULAR", 5000),
                        (u"OPERADORA", 5000),
                        (u"TIPO DE BECA", 5000)
                    ]
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], fuentecabecera)
                        ws.col(col_num).width = columns[col_num][1]

                    row_num = 6

                    # if tipobeca == 0:
                    #     tipos = [18, 19]
                    # else:
                    #     tipos = [tipobeca]

                    becas1 = BecaAsignacion.objects.filter(cargadocumento=True,
                                                           solicitud__inscripcion__persona__personadocumentopersonal__estadocedula=2,
                                                           solicitud__inscripcion__persona__personadocumentopersonal__estadocedularepresentantesol=2,
                                                           solicitud__inscripcion__persona__personadocumentopersonal__estadopapeleta=2,
                                                           solicitud__inscripcion__persona__personadocumentopersonal__estadopapeletarepresentantesol=2,
                                                           solicitud__inscripcion__persona__personadocumentopersonal__fecha_modificacion__range=(
                                                           desde, hasta),
                                                           status=True, solicitud__periodo=periodo, tipo=1).exclude(
                        solicitud__becatipo_id=16)

                    becas2 = BecaAsignacion.objects.filter(cargadocumento=True,
                                                           solicitud__inscripcion__persona__personadocumentopersonal__estadocedula=2,
                                                           solicitud__inscripcion__persona__personadocumentopersonal__estadocedularepresentantesol=2,
                                                           solicitud__inscripcion__persona__personadocumentopersonal__estadopapeleta=2,
                                                           solicitud__inscripcion__persona__personadocumentopersonal__estadopapeletarepresentantesol=2,
                                                           solicitud__inscripcion__persona__personadocumentopersonal__estadoactagrado=2,
                                                           solicitud__inscripcion__persona__personadocumentopersonal__fecha_modificacion__range=(
                                                           desde, hasta),
                                                           status=True, solicitud__periodo=periodo,
                                                           solicitud__becatipo_id=16, tipo=1)

                    becas = becas1 | becas2
                    becas = becas.order_by('-solicitud_id')

                    # becas = BecaAsignacion.objects.filter(
                    #                                  status=True, solicitud__periodo=periodo).order_by('-solicitud_id')

                    for s in becas:
                        row_num += 1
                        print(row_num)
                        p = Persona.objects.get(pk=s.solicitud.inscripcion.persona_id)

                        umat = s.solicitud.inscripcion.matricula_periodo_actual(periodo)[0]

                        ws.write(row_num, 0, str(s.solicitud.id), fuentenumeroentero)

                        d = p.documentos_personales()

                        # ws.write(row_num, 1, str(s.fecha_creacion)[:10], fuentefecha)
                        ws.write(row_num, 1, str(d.fecha_modificacion)[:10], fuentefecha)
                        ws.write(row_num, 2, s.montobeneficio, fuentemoneda)
                        ws.write(row_num, 3, p.nombre_completo_inverso(), fuentenormal)
                        ws.write(row_num, 4, p.identificacion(), fuentenormal)

                        ws.write(row_num, 5, str(umat.nivel.coordinacion()), fuentenormal)
                        ws.write(row_num, 6, str(umat.inscripcion.carrera), fuentenormal)
                        ws.write(row_num, 7, str(umat.inscripcion.modalidad), fuentenormal)
                        ws.write(row_num, 8, umat.nivelmalla.nombre, fuentenormal)
                        ws.write(row_num, 9, str(p.provincia) if p.provincia else '', fuentenormal)
                        ws.write(row_num, 10, str(p.canton) if p.canton else '', fuentenormal)
                        ws.write(row_num, 11, str(p.parroquia) if p.parroquia else '', fuentenormal)
                        ws.write(row_num, 12, p.direccion_corta().upper(), fuentenormal)
                        ws.write(row_num, 13, p.referencia.upper(), fuentenormal)
                        ws.write(row_num, 14, p.sector.upper(), fuentenormal)
                        ws.write(row_num, 15, p.num_direccion, fuentenormal)
                        ws.write(row_num, 16, p.email, fuentenormal)
                        ws.write(row_num, 17, p.emailinst, fuentenormal)
                        ws.write(row_num, 18, p.telefono_conv, fuentenormal)
                        ws.write(row_num, 19, p.telefono, fuentenormal)
                        ws.write(row_num, 20, p.get_tipocelular_display() if p.tipocelular else '', fuentenormal)
                        ws.write(row_num, 21, s.solicitud.becatipo.nombrecorto, fuentenormal)

                    wb.save(response)
                    return response
                except Exception as ex:
                    print("Error...")
                    pass

            elif action == 'becadosdocumentacionpendiente':
                try:
                    return generar_reporte_pendientes_cargar_requisitos(request, periodo)
                except Exception as ex:
                    messages.error(request, str(ex))

            elif action == 'becadosdocumentacionrechazada':
                try:
                    __author__ = 'Unemi'
                    titulo = easyxf(
                        'font: name Times New Roman, color-index black, bold on , height 350; alignment: horiz centre')
                    titulo2 = easyxf(
                        'font: name Times New Roman, color-index black, bold on , height 250; alignment: horiz centre')
                    fuentecabecera = easyxf(
                        'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
                    fuentenormal = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
                    fuentemoneda = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right',
                        num_format_str=' "$" #,##0.00')
                    fuentefecha = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz center',
                        num_format_str='yyyy-mm-dd')
                    fuentenumerodecimal = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right',
                        num_format_str='#,##0.00')
                    fuentenumeroentero = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right')

                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('Becados')
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=becados_docurechazada_' + random.randint(1,
                                                                                                                     10000).__str__() + '.xls'

                    ws.write_merge(0, 0, 0, 12, 'UNIVERSIDAD ESTATAL DE MILAGRO', titulo)
                    ws.write_merge(1, 1, 0, 12, 'DIRECCIÓN DE BIENESTAR UNIVERSITARIO', titulo2)
                    ws.write_merge(2, 2, 0, 12, 'LISTADO DE BECADOS - DOCUMENTACIÓN RECHAZADA Y PENDIENTE DE CARGAR',
                                   titulo2)

                    row_num = 4
                    columns = [
                        (u"# SOLICITUD", 3000),
                        (u"FECHA BECA", 3000),
                        (u"MONTO BECA", 3000),
                        (u"NOMBRES COMPLETOS", 10000),
                        (u"IDENTIFICACIÓN", 5000),
                        (u"FACULTAD", 5000),
                        (u"CARRERA", 5000),
                        (u"MODALIDAD", 5000),
                        (u"NIVEL", 5000),
                        (u"PROVINCIA", 5000),
                        (u"CANTÓN", 5000),
                        (u"PARROQUIA", 5000),
                        (u"DIRECCIÓN", 15000),
                        (u"REFERENCIA", 15000),
                        (u"SECTOR", 15000),
                        (u"# CASA", 5000),
                        (u"# CORREO PERSONAL", 8000),
                        (u"# CORREO UNEMI", 8000),
                        (u"TELÉFONO", 5000),
                        (u"CELULAR", 5000),
                        (u"OPERADORA", 5000),
                        (u"MOTIVO RECHAZO", 20000)
                    ]
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], fuentecabecera)
                        ws.col(col_num).width = columns[col_num][1]

                    row_num = 4

                    becas1 = BecaAsignacion.objects.filter(
                        Q(solicitud__inscripcion__persona__personadocumentopersonal__estadocedula=3) |
                        Q(solicitud__inscripcion__persona__personadocumentopersonal__estadopapeleta=3) |
                        Q(solicitud__inscripcion__persona__personadocumentopersonal__estadocedularepresentantesol=3) |
                        Q(solicitud__inscripcion__persona__personadocumentopersonal__estadopapeletarepresentantesol=3),
                        status=True, solicitud__periodo=periodo, tipo=1, cargadocumento=True).exclude(
                        solicitud__becatipo_id=16)

                    becas2 = BecaAsignacion.objects.filter(
                        Q(solicitud__inscripcion__persona__personadocumentopersonal__estadocedula=3) |
                        Q(solicitud__inscripcion__persona__personadocumentopersonal__estadopapeleta=3) |
                        Q(solicitud__inscripcion__persona__personadocumentopersonal__estadocedularepresentantesol=3) |
                        Q(solicitud__inscripcion__persona__personadocumentopersonal__estadopapeletarepresentantesol=3) |
                        Q(solicitud__inscripcion__persona__personadocumentopersonal__estadoactagrado=3),
                        status=True, solicitud__periodo=periodo, tipo=1, solicitud__becatipo_id=16, cargadocumento=True)

                    becas = becas1 | becas2

                    becas = becas.order_by('-solicitud_id')

                    # becas = BecaAsignacion.objects.filter(Q(solicitud__inscripcion__persona__personadocumentopersonal__estadocedula=3)|
                    #                                     Q(solicitud__inscripcion__persona__personadocumentopersonal__estadopapeleta=3)|
                    #                                     Q(solicitud__inscripcion__persona__personadocumentopersonal__estadocedularepresentantesol=3)|
                    #                                     Q(solicitud__inscripcion__persona__personadocumentopersonal__estadopapeletarepresentantesol=3),
                    #                                     status=True, solicitud__periodo=periodo).order_by('-solicitud_id')

                    for s in becas:
                        row_num += 1
                        p = Persona.objects.get(pk=s.solicitud.inscripcion.persona_id)

                        umat = s.solicitud.inscripcion.matricula_periodo_actual(periodo)[0]

                        ws.write(row_num, 0, str(s.solicitud.id), fuentenumeroentero)
                        ws.write(row_num, 1, str(s.fecha_creacion)[:10], fuentefecha)
                        ws.write(row_num, 2, s.montobeneficio, fuentemoneda)
                        ws.write(row_num, 3, p.nombre_completo_inverso(), fuentenormal)
                        ws.write(row_num, 4, p.identificacion(), fuentenormal)

                        ws.write(row_num, 5, str(umat.nivel.coordinacion()), fuentenormal)
                        ws.write(row_num, 6, str(umat.inscripcion.carrera), fuentenormal)
                        ws.write(row_num, 7, str(umat.inscripcion.modalidad), fuentenormal)
                        ws.write(row_num, 8, umat.nivelmalla.nombre, fuentenormal)
                        ws.write(row_num, 9, str(p.provincia) if p.provincia else '', fuentenormal)
                        ws.write(row_num, 10, str(p.canton) if p.canton else '', fuentenormal)
                        ws.write(row_num, 11, str(p.parroquia) if p.parroquia else '', fuentenormal)
                        ws.write(row_num, 12, p.direccion_corta().upper(), fuentenormal)
                        ws.write(row_num, 13, p.referencia.upper(), fuentenormal)
                        ws.write(row_num, 14, p.sector.upper(), fuentenormal)
                        ws.write(row_num, 15, p.num_direccion, fuentenormal)
                        ws.write(row_num, 16, p.email, fuentenormal)
                        ws.write(row_num, 17, p.emailinst, fuentenormal)
                        ws.write(row_num, 18, p.telefono_conv, fuentenormal)
                        ws.write(row_num, 19, p.telefono, fuentenormal)
                        ws.write(row_num, 20, p.get_tipocelular_display() if p.tipocelular else '', fuentenormal)
                        ws.write(row_num, 21, p.documentos_personales().observacion, fuentenormal)
                    wb.save(response)

                    return response
                except Exception as ex:
                    print("Error...")
                    pass

            elif action == 'becadoscontratopendiente':
                try:
                    __author__ = 'Unemi'
                    titulo = easyxf(
                        'font: name Times New Roman, color-index black, bold on , height 350; alignment: horiz centre')
                    titulo2 = easyxf(
                        'font: name Times New Roman, color-index black, bold on , height 250; alignment: horiz centre')
                    fuentecabecera = easyxf(
                        'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
                    fuentenormal = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
                    fuentemoneda = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right',
                        num_format_str=' "$" #,##0.00')
                    fuentefecha = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz center',
                        num_format_str='yyyy-mm-dd')
                    fuentenumerodecimal = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right',
                        num_format_str='#,##0.00')
                    fuentenumeroentero = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right')

                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('Becados')
                    response = HttpResponse(content_type="application/ms-excel")
                    response[
                        'Content-Disposition'] = 'attachment; filename=becados_contratopendiente_' + random.randint(1,
                                                                                                                    10000).__str__() + '.xls'

                    ws.write_merge(0, 0, 0, 12, 'UNIVERSIDAD ESTATAL DE MILAGRO', titulo)
                    ws.write_merge(1, 1, 0, 12, 'DIRECCIÓN DE BIENESTAR UNIVERSITARIO', titulo2)
                    ws.write_merge(2, 2, 0, 12, 'LISTADO DE BECADOS - CONTRATOS PENDIENTES DE GENERAR', titulo2)

                    row_num = 4
                    columns = [
                        (u"# SOLICITUD", 3000),
                        (u"FECHA BECA", 3000),
                        (u"MONTO BECA", 3000),
                        (u"NOMBRES COMPLETOS", 10000),
                        (u"IDENTIFICACIÓN", 5000),
                        (u"FACULTAD", 5000),
                        (u"CARRERA", 5000),
                        (u"MODALIDAD", 5000),
                        (u"NIVEL", 5000),
                        (u"PROVINCIA", 5000),
                        (u"CANTÓN", 5000),
                        (u"PARROQUIA", 5000),
                        (u"DIRECCIÓN", 15000),
                        (u"REFERENCIA", 15000),
                        (u"SECTOR", 15000),
                        (u"# CASA", 5000),
                        (u"# CORREO PERSONAL", 8000),
                        (u"# CORREO UNEMI", 8000),
                        (u"TELÉFONO", 5000),
                        (u"CELULAR", 5000),
                        (u"OPERADORA", 5000)
                    ]
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], fuentecabecera)
                        ws.col(col_num).width = columns[col_num][1]

                    row_num = 4

                    becas1 = BecaAsignacion.objects.filter(
                        solicitud__inscripcion__persona__personadocumentopersonal__status=True,
                        solicitud__inscripcion__persona__personadocumentopersonal__estadocedula=2,
                        solicitud__inscripcion__persona__personadocumentopersonal__estadopapeleta=2,
                        solicitud__inscripcion__persona__personadocumentopersonal__estadocedularepresentantesol=2,
                        solicitud__inscripcion__persona__personadocumentopersonal__estadopapeletarepresentantesol=2,
                        numerocontrato__isnull=True,
                        status=True, solicitud__periodo=periodo, tipo=1, cargadocumento=True).exclude(
                        solicitud__becatipo_id=16)

                    becas2 = BecaAsignacion.objects.filter(
                        solicitud__inscripcion__persona__personadocumentopersonal__status=True,
                        solicitud__inscripcion__persona__personadocumentopersonal__estadocedula=2,
                        solicitud__inscripcion__persona__personadocumentopersonal__estadopapeleta=2,
                        solicitud__inscripcion__persona__personadocumentopersonal__estadocedularepresentantesol=2,
                        solicitud__inscripcion__persona__personadocumentopersonal__estadopapeletarepresentantesol=2,
                        solicitud__inscripcion__persona__personadocumentopersonal__estadoactagrado=2,
                        numerocontrato__isnull=True,
                        status=True, solicitud__periodo=periodo, tipo=1, solicitud__becatipo_id=16, cargadocumento=True)

                    becas3 = BecaAsignacion.objects.filter(
                        solicitud__inscripcion__persona__personadocumentopersonal__status=True,
                        numerocontrato__isnull=True,
                        status=True, solicitud__periodo=periodo, tipo=1, cargadocumento=False)

                    becas = becas1 | becas2 | becas3

                    becas = becas.order_by('-solicitud_id')

                    # becas = BecaAsignacion.objects.filter(solicitud__inscripcion__persona__personadocumentopersonal__status=True,
                    #                                         solicitud__inscripcion__persona__personadocumentopersonal__estadocedula=2,
                    #                                         solicitud__inscripcion__persona__personadocumentopersonal__estadopapeleta=2,
                    #                                         solicitud__inscripcion__persona__personadocumentopersonal__estadocedularepresentantesol=2,
                    #                                         solicitud__inscripcion__persona__personadocumentopersonal__estadopapeletarepresentantesol=2,
                    #                                         numerocontrato__isnull=True,
                    #                                         status=True, solicitud__periodo=periodo).order_by('-solicitud_id')

                    for s in becas:
                        row_num += 1
                        p = Persona.objects.get(pk=s.solicitud.inscripcion.persona_id)

                        umat = s.solicitud.inscripcion.matricula_periodo_actual(periodo)[0]

                        ws.write(row_num, 0, str(s.solicitud.id), fuentenumeroentero)
                        ws.write(row_num, 1, str(s.fecha_creacion)[:10], fuentefecha)
                        ws.write(row_num, 2, s.montobeneficio, fuentemoneda)
                        ws.write(row_num, 3, p.nombre_completo_inverso(), fuentenormal)
                        ws.write(row_num, 4, p.identificacion(), fuentenormal)

                        ws.write(row_num, 5, str(umat.nivel.coordinacion()), fuentenormal)
                        ws.write(row_num, 6, str(umat.inscripcion.carrera), fuentenormal)
                        ws.write(row_num, 7, str(umat.inscripcion.modalidad), fuentenormal)
                        ws.write(row_num, 8, umat.nivelmalla.nombre, fuentenormal)
                        ws.write(row_num, 9, str(p.provincia) if p.provincia else '', fuentenormal)
                        ws.write(row_num, 10, str(p.canton) if p.canton else '', fuentenormal)
                        ws.write(row_num, 11, str(p.parroquia) if p.parroquia else '', fuentenormal)
                        ws.write(row_num, 12, p.direccion_corta().upper(), fuentenormal)
                        ws.write(row_num, 13, p.referencia.upper(), fuentenormal)
                        ws.write(row_num, 14, p.sector.upper(), fuentenormal)
                        ws.write(row_num, 15, p.num_direccion, fuentenormal)
                        ws.write(row_num, 16, p.email, fuentenormal)
                        ws.write(row_num, 17, p.emailinst, fuentenormal)
                        ws.write(row_num, 18, p.telefono_conv, fuentenormal)
                        ws.write(row_num, 19, p.telefono, fuentenormal)
                        ws.write(row_num, 20, p.get_tipocelular_display() if p.tipocelular else '', fuentenormal)

                    wb.save(response)
                    return response
                except Exception as ex:
                    print("Error...")
                    pass

            elif action == 'becadoscontratopendientecargar':
                try:
                    __author__ = 'Unemi'
                    titulo = easyxf(
                        'font: name Times New Roman, color-index black, bold on , height 350; alignment: horiz centre')
                    titulo2 = easyxf(
                        'font: name Times New Roman, color-index black, bold on , height 250; alignment: horiz centre')
                    fuentecabecera = easyxf(
                        'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
                    fuentenormal = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
                    fuentemoneda = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right',
                        num_format_str=' "$" #,##0.00')
                    fuentefecha = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz center',
                        num_format_str='yyyy-mm-dd')
                    fuentenumerodecimal = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right',
                        num_format_str='#,##0.00')
                    fuentenumeroentero = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right')

                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('Becados')
                    response = HttpResponse(content_type="application/ms-excel")
                    response[
                        'Content-Disposition'] = 'attachment; filename=becados_contratofirmadopendiente_' + random.randint(
                        1, 10000).__str__() + '.xls'

                    ws.write_merge(0, 0, 0, 12, 'UNIVERSIDAD ESTATAL DE MILAGRO', titulo)
                    ws.write_merge(1, 1, 0, 12, 'DIRECCIÓN DE BIENESTAR UNIVERSITARIO', titulo2)
                    ws.write_merge(2, 2, 0, 12, 'LISTADO DE BECADOS - CONTRATOS FIRMADOS PENDIENTES DE CARGAR', titulo2)

                    row_num = 4
                    columns = [
                        (u"# SOLICITUD", 3000),
                        (u"FECHA BECA", 3000),
                        (u"MONTO BECA", 3000),
                        (u"NOMBRES COMPLETOS", 10000),
                        (u"IDENTIFICACIÓN", 5000),
                        (u"FACULTAD", 5000),
                        (u"CARRERA", 5000),
                        (u"MODALIDAD", 5000),
                        (u"NIVEL", 5000),
                        (u"PROVINCIA", 5000),
                        (u"CANTÓN", 5000),
                        (u"PARROQUIA", 5000),
                        (u"DIRECCIÓN", 15000),
                        (u"REFERENCIA", 15000),
                        (u"SECTOR", 15000),
                        (u"# CASA", 5000),
                        (u"# CORREO PERSONAL", 8000),
                        (u"# CORREO UNEMI", 8000),
                        (u"TELÉFONO", 5000),
                        (u"CELULAR", 5000),
                        (u"OPERADORA", 5000)
                    ]
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], fuentecabecera)
                        ws.col(col_num).width = columns[col_num][1]

                    row_num = 4

                    becas = BecaAsignacion.objects.filter(Q(archivocontrato='') | Q(archivocontrato__isnull=True),
                                                          numerocontrato__isnull=False,
                                                          status=True, solicitud__periodo=periodo).order_by(
                        '-solicitud_id')

                    for s in becas:
                        row_num += 1
                        p = Persona.objects.get(pk=s.solicitud.inscripcion.persona_id)

                        umat = s.solicitud.inscripcion.matricula_periodo_actual(periodo)[0]

                        ws.write(row_num, 0, str(s.solicitud.id), fuentenumeroentero)
                        ws.write(row_num, 1, str(s.fecha_creacion)[:10], fuentefecha)
                        ws.write(row_num, 2, s.montobeneficio, fuentemoneda)
                        ws.write(row_num, 3, p.nombre_completo_inverso(), fuentenormal)
                        ws.write(row_num, 4, p.identificacion(), fuentenormal)

                        ws.write(row_num, 5, str(umat.nivel.coordinacion()), fuentenormal)
                        ws.write(row_num, 6, str(umat.inscripcion.carrera), fuentenormal)
                        ws.write(row_num, 7, str(umat.inscripcion.modalidad), fuentenormal)
                        ws.write(row_num, 8, umat.nivelmalla.nombre, fuentenormal)
                        ws.write(row_num, 9, str(p.provincia) if p.provincia else '', fuentenormal)
                        ws.write(row_num, 10, str(p.canton) if p.canton else '', fuentenormal)
                        ws.write(row_num, 11, str(p.parroquia) if p.parroquia else '', fuentenormal)
                        ws.write(row_num, 12, p.direccion_corta().upper(), fuentenormal)
                        ws.write(row_num, 13, p.referencia.upper(), fuentenormal)
                        ws.write(row_num, 14, p.sector.upper(), fuentenormal)
                        ws.write(row_num, 15, p.num_direccion, fuentenormal)
                        ws.write(row_num, 16, p.email, fuentenormal)
                        ws.write(row_num, 17, p.emailinst, fuentenormal)
                        ws.write(row_num, 18, p.telefono_conv, fuentenormal)
                        ws.write(row_num, 19, p.telefono, fuentenormal)
                        ws.write(row_num, 20, p.get_tipocelular_display() if p.tipocelular else '', fuentenormal)

                    wb.save(response)
                    return response
                except Exception as ex:
                    print("Error...")
                    pass

            elif action == 'becadoscontratorechazado':
                try:
                    __author__ = 'Unemi'
                    titulo = easyxf(
                        'font: name Times New Roman, color-index black, bold on , height 350; alignment: horiz centre')
                    titulo2 = easyxf(
                        'font: name Times New Roman, color-index black, bold on , height 250; alignment: horiz centre')
                    fuentecabecera = easyxf(
                        'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
                    fuentenormal = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
                    fuentemoneda = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right',
                        num_format_str=' "$" #,##0.00')
                    fuentefecha = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz center',
                        num_format_str='yyyy-mm-dd')
                    fuentenumerodecimal = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right',
                        num_format_str='#,##0.00')
                    fuentenumeroentero = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right')

                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('Becados')
                    response = HttpResponse(content_type="application/ms-excel")
                    response[
                        'Content-Disposition'] = 'attachment; filename=becados_contrato_rechazado_' + random.randint(1,
                                                                                                                     10000).__str__() + '.xls'

                    ws.write_merge(0, 0, 0, 12, 'UNIVERSIDAD ESTATAL DE MILAGRO', titulo)
                    ws.write_merge(1, 1, 0, 12, 'DIRECCIÓN DE BIENESTAR UNIVERSITARIO', titulo2)
                    ws.write_merge(2, 2, 0, 12, 'LISTADO DE BECADOS - CONTRATOS RECHAZADOS Y PENDIENTES DE CARGAR',
                                   titulo2)

                    row_num = 4
                    columns = [
                        (u"# SOLICITUD", 3000),
                        (u"FECHA BECA", 3000),
                        (u"MONTO BECA", 3000),
                        (u"NOMBRES COMPLETOS", 10000),
                        (u"IDENTIFICACIÓN", 5000),
                        (u"FACULTAD", 5000),
                        (u"CARRERA", 5000),
                        (u"MODALIDAD", 5000),
                        (u"NIVEL", 5000),
                        (u"PROVINCIA", 5000),
                        (u"CANTÓN", 5000),
                        (u"PARROQUIA", 5000),
                        (u"DIRECCIÓN", 15000),
                        (u"REFERENCIA", 15000),
                        (u"SECTOR", 15000),
                        (u"# CASA", 5000),
                        (u"# CORREO PERSONAL", 8000),
                        (u"# CORREO UNEMI", 8000),
                        (u"TELÉFONO", 5000),
                        (u"CELULAR", 5000),
                        (u"OPERADORA", 5000),
                        (u"MOTIVO RECHAZO", 20000)
                    ]
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], fuentecabecera)
                        ws.col(col_num).width = columns[col_num][1]

                    row_num = 4

                    becas = BecaAsignacion.objects.filter(~Q(archivocontrato=''),
                                                          archivocontrato__isnull=False,
                                                          numerocontrato__isnull=False,
                                                          estadorevisioncontrato=3,
                                                          status=True, solicitud__periodo=periodo).order_by(
                        '-solicitud_id')

                    for s in becas:
                        row_num += 1
                        p = Persona.objects.get(pk=s.solicitud.inscripcion.persona_id)

                        umat = s.solicitud.inscripcion.matricula_periodo_actual(periodo)[0]

                        ws.write(row_num, 0, str(s.solicitud.id), fuentenumeroentero)
                        ws.write(row_num, 1, str(s.fecha_creacion)[:10], fuentefecha)
                        ws.write(row_num, 2, s.montobeneficio, fuentemoneda)
                        ws.write(row_num, 3, p.nombre_completo_inverso(), fuentenormal)
                        ws.write(row_num, 4, p.identificacion(), fuentenormal)

                        ws.write(row_num, 5, str(umat.nivel.coordinacion()), fuentenormal)
                        ws.write(row_num, 6, str(umat.inscripcion.carrera), fuentenormal)
                        ws.write(row_num, 7, str(umat.inscripcion.modalidad), fuentenormal)
                        ws.write(row_num, 8, umat.nivelmalla.nombre, fuentenormal)
                        ws.write(row_num, 9, str(p.provincia) if p.provincia else '', fuentenormal)
                        ws.write(row_num, 10, str(p.canton) if p.canton else '', fuentenormal)
                        ws.write(row_num, 11, str(p.parroquia) if p.parroquia else '', fuentenormal)
                        ws.write(row_num, 12, p.direccion_corta().upper(), fuentenormal)
                        ws.write(row_num, 13, p.referencia.upper(), fuentenormal)
                        ws.write(row_num, 14, p.sector.upper(), fuentenormal)
                        ws.write(row_num, 15, p.num_direccion, fuentenormal)
                        ws.write(row_num, 16, p.email, fuentenormal)
                        ws.write(row_num, 17, p.emailinst, fuentenormal)
                        ws.write(row_num, 18, p.telefono_conv, fuentenormal)
                        ws.write(row_num, 19, p.telefono, fuentenormal)
                        ws.write(row_num, 20, p.get_tipocelular_display() if p.tipocelular else '', fuentenormal)
                        ws.write(row_num, 21, s.observacion, fuentenormal)

                    wb.save(response)
                    return response
                except Exception as ex:
                    print("Error...")
                    pass

            elif action == 'becadosmatrizsiiescaces':
                try:
                    notifi = Notificacion(cuerpo='Generación de reporte de excel en progreso',
                                          titulo='Reporte de becas (Matriz SIIES-CACES)',
                                          destinatario=persona,
                                          url='',
                                          prioridad=1, app_label='SGA',
                                          fecha_hora_visible=datetime.now() + timedelta(days=1), tipo=2,
                                          en_proceso=True)
                    notifi.save(request)
                    reportebecadosmatrizsiiescaces(request=request, notiid=notifi.id, periodo=periodo).start()
                    return JsonResponse({"result": True,
                                         "mensaje": u"El reporte de becas (Matriz SIIES-CACES) se está realizando. Verifique su apartado de notificaciones después de unos minutos.",
                                         "btn_notificaciones": traerNotificaciones(request, data, persona)})
                except Exception as ex:
                    print("Error...")
                    pass

            elif action == 'reportesolicitudpagoexcel':
                try:
                    __author__ = 'Unemi'
                    titulo = easyxf(
                        'font: name Times New Roman, color-index black, bold on , height 350; alignment: horiz centre')
                    titulo2 = easyxf(
                        'font: name Times New Roman, color-index black, bold on , height 250; alignment: horiz centre')
                    fuentecabecera = easyxf(
                        'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
                    fuentenormal = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
                    fuentemoneda = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right',
                        num_format_str=' "$" #,##0.00')
                    fuentefecha = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz center',
                        num_format_str='yyyy-mm-dd')
                    fuentenumerodecimal = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right',
                        num_format_str='#,##0.00')
                    fuentenumeroentero = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right')

                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('Becados')
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=becados_solicitud_pago_' + random.randint(1,
                                                                                                                      10000).__str__() + '.xls'

                    ws.write_merge(0, 0, 0, 7, 'UNIVERSIDAD ESTATAL DE MILAGRO', titulo)
                    ws.write_merge(1, 1, 0, 7, 'DIRECCIÓN DE BIENESTAR UNIVERSITARIO', titulo2)
                    ws.write_merge(2, 2, 0, 7, 'REPORTE SOLICITUD DE PAGO PARA BECADOS', titulo2)

                    row_num = 4
                    columns = [
                        (u"# SOLICITUD", 3000),
                        (u"FECHA BECA", 3000),
                        (u"IDENTIFICACIÓN", 5000),
                        (u"NOMBRES COMPLETOS", 10000),
                        (u"N° CUENTA", 5000),
                        (u"BANCO", 10000),
                        (u"MONTO BECA", 3000)
                    ]
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], fuentecabecera)
                        ws.col(col_num).width = columns[col_num][1]

                    row_num = 4

                    becas = BecaAsignacion.objects.filter(
                        solicitud__inscripcion__persona__personadocumentopersonal__estadocedula=2,
                        solicitud__inscripcion__persona__personadocumentopersonal__estadopapeleta=2,
                        solicitud__inscripcion__persona__personadocumentopersonal__estadocedularepresentantesol=2,
                        solicitud__inscripcion__persona__personadocumentopersonal__estadopapeletarepresentantesol=2,
                        estadorevisioncontrato=2,
                        solicitud__inscripcion__persona__cuentabancariapersona__estadorevision=2,
                        status=True, solicitud__periodo=periodo).order_by('-solicitud_id')

                    for s in becas:
                        row_num += 1
                        p = Persona.objects.get(pk=s.solicitud.inscripcion.persona_id)

                        ws.write(row_num, 0, str(s.solicitud.id), fuentenumeroentero)
                        ws.write(row_num, 1, str(s.fecha_creacion)[:10], fuentefecha)
                        ws.write(row_num, 2, p.identificacion(), fuentenormal)
                        ws.write(row_num, 3, p.nombre_completo_inverso(), fuentenormal)
                        ws.write(row_num, 4, p.cuentabancaria().numero, fuentenormal)
                        ws.write(row_num, 5, p.cuentabancaria().banco.nombre, fuentenormal)
                        ws.write(row_num, 6, s.montobeneficio, fuentemoneda)

                    wb.save(response)

                    return response
                except Exception as ex:
                    print("Error...")
                    pass

            elif action == 'mostrarevidencias':
                try:
                    data = {}
                    beca = BecaAsignacion.objects.get(pk=int(request.GET['id']))
                    data['beca'] = beca
                    template = get_template("adm_becas/mostrarevidencianetrega.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content})
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

            elif action == 'listasolicitudpago':
                try:
                    data = {}
                    solicitudes = SolicitudPagoBeca.objects.filter(status=True).order_by('-id')
                    data['solicitudes'] = solicitudes
                    template = get_template("adm_becas/consultasolicitudpago.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content})
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

            elif action == 'subirevidenciasentrega':
                try:
                    beca = BecaAsignacion.objects.get(pk=int(request.GET['id']))
                    obligatorio = False
                    form = BecaAsignacionSubirEvidenciaForm()
                    if not beca.archivocontrato:
                        obligatorio = True

                    data['title'] = u'Subir Evidencias - Contrato # ' + str(beca.numerocontrato).zfill(4)
                    data['id'] = int(request.GET['id'])
                    data['form'] = form
                    data['obligatorio'] = obligatorio

                    template = get_template("adm_becas/subirevidenciaentrega.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content, 'title': data['title']})
                except Exception as ex:
                    pass

            elif action == 'listatipobeca':
                try:
                    data['title'] = u'Gestión de Tipo de Becas'
                    search = None
                    ids = None
                    tipobecas = BecaTipo.objects.filter(status=True).order_by('nombre', 'vigente')
                    if 's' in request.GET:
                        search = request.GET['s']
                        if search.isdigit():
                            tipobecas = tipobecas.filter(pk=search)
                        else:
                            tipobecas = tipobecas.filter(Q(nombre__icontains=search))
                    if 'id' in request.GET:
                        ids = request.GET['id']
                        tipobecas = tipobecas.filter(id=ids)
                    vigente = 0
                    if 'vigente' in request.GET and int(request.GET['vigente']) > 0:
                        vigente = int(request.GET['vigente'])
                        tipobecas = tipobecas.filter(vigente=int(request.GET['vigente']) == 1)
                    paging = MiPaginador(tipobecas, 25)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['page'] = page
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['listatipobeca'] = page.object_list
                    data['vigente'] = vigente
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    return render(request, "adm_becas/listatipobeca.html", data)
                except Exception as ex:
                    pass

            elif action == 'listadobecados':
                try:
                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('becados')
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=becados' + random.randint(1,
                                                                                                      10000).__str__() + '.xls'
                    columns = [
                        (u"CODIGO_IES", 2000),
                        (u"CODIGO_CARRERA", 2000),
                        (u"TIPO_IDENTIFICACION", 2000),
                        (u"IDENTIFICACION", 2000),
                        (u"CODIGO_BECA", 2000),
                        (u"ANIO", 2000),
                        (u"FECHA_INICIO_PERIODO", 2000),
                        (u"FECHA_FIN_PERIODO", 2000),
                        (u"TIPO_AYUDA", 2000),
                        (u"MOTIVO_BECA", 12000),
                        (u"OTRO_MOTIVO", 2000),
                        (u"MONTO_RECIBIDO", 2000),
                        (u"PORCENTAJE_VALOR_ARANCEL", 2000),
                        (u"PORCENTAJE_MANUTENCION", 2000),
                        (u"TIPO_FINANCIAMIENTO", 12000),
                        (u"NOMBRE", 12000),
                        (u"CARRERA", 12000),
                        (u"DIRECCION", 12000),
                        (u"GRUPO SOCIOECONOMICO", 12000),
                        (u"SEXO", 12000),
                        (u"PROMEDIO", 12000)

                    ]
                    row_num = 0
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], font_style)
                        ws.col(col_num).width = columns[col_num][1]
                    date_format = xlwt.XFStyle()
                    date_format.num_format_str = 'yyyy/mm/dd'
                    row_num = 1
                    resultados = BecaAsignacion.objects.filter(status=True, solicitud__periodo=periodo,
                                                               solicitud__status=True)
                    for x in resultados:
                        campo1 = '1024'
                        if (x.solicitud.inscripcion.mi_malla()):
                            campo2 = x.solicitud.inscripcion.mi_malla().codigo
                        else:
                            campo2 = ''
                        campo3 = str(x.solicitud.inscripcion.persona.tipo_identificacion_completo())
                        campo4 = x.solicitud.inscripcion.persona.identificacion()
                        if x.solicitud.inscripcion.matricula_periodo(periodo):
                            campo5 = str(x.solicitud.inscripcion.matricula_periodo(periodo).id) + "-" + str(
                                x.solicitud.becatipo_id)
                        else:
                            campo5 = ''
                        campo6 = periodo.anio
                        campo7 = str(periodo.inicio)
                        campo8 = str(periodo.fin)
                        campo9 = 'BECA COMPLETA'
                        campo10 = x.solicitud.becatipo.nombre
                        campo11 = ''
                        campo12 = str(x.montobeneficio)
                        campo13 = ''
                        campo14 = ''
                        campo15 = 'RECURSOS FISCALES'
                        campo16 = x.solicitud.inscripcion.persona.nombre_completo_inverso()
                        campo17 = x.solicitud.inscripcion.carerra
                        campo18 = x.solicitud.inscripcion.persona.direccion_completa()
                        campo19 = 'BAJO'
                        campo20 = x.solicitud.inscripcion.persona.sexo
                        campo21 = 'PROMEDIO'
                        ws.write(row_num, 0, campo1, font_style2)
                        ws.write(row_num, 1, campo2, font_style2)
                        ws.write(row_num, 2, campo3, font_style2)
                        ws.write(row_num, 3, campo4, font_style2)
                        ws.write(row_num, 4, campo5, font_style2)
                        ws.write(row_num, 5, campo6, font_style2)
                        ws.write(row_num, 6, campo7, font_style2)
                        ws.write(row_num, 7, campo8, font_style2)
                        ws.write(row_num, 8, campo9, font_style2)
                        ws.write(row_num, 9, campo10, font_style2)
                        ws.write(row_num, 10, campo11, font_style2)
                        ws.write(row_num, 11, campo12, font_style2)
                        ws.write(row_num, 12, campo13, font_style2)
                        ws.write(row_num, 13, campo14, font_style2)
                        ws.write(row_num, 14, campo15, font_style2)
                        ws.write(row_num, 15, campo16, font_style2)
                        ws.write(row_num, 16, campo17, font_style2)
                        ws.write(row_num, 17, campo18, font_style2)
                        ws.write(row_num, 18, campo19, font_style2)
                        ws.write(row_num, 19, campo20, font_style2)
                        row_num += 1
                    wb.save(response)
                    return response
                except Exception as ex:
                    pass

            elif action == 'addtipobeca':
                try:
                    data['title'] = u'Agregar tipo Beca'
                    form = BecaTipoForm()
                    data['form'] = form
                    return render(request, "adm_becas/addtipobeca.html", data)
                except Exception as ex:
                    pass

            elif action == 'edititipobeca':
                try:
                    data['title'] = u'Editar Tipo Beca'
                    data['tipobeca'] = tipobeca = BecaTipo.objects.get(pk=int(request.GET['id']))
                    form = BecaTipoForm(initial={'nombre': tipobeca.nombre,
                                                 'nombrecorto': tipobeca.nombrecorto,
                                                 'minimo_asistencia': tipobeca.minimo_asistencia,
                                                 'minimo_promedio': tipobeca.minimo_promedio,
                                                 'numemejores': tipobeca.numemejores,
                                                 'vigente': tipobeca.vigente,
                                                 'nombrecaces': tipobeca.nombrecaces})
                    data['form'] = form
                    return render(request, "adm_becas/edittipobeca.html", data)
                except Exception as ex:
                    pass

            elif action == 'deletipobeca':
                try:
                    data['title'] = u'Eliminar Tipo Beca'
                    data['tipobeca'] = BecaTipo.objects.get(pk=int(request.GET['id']), status=True,
                                                            vigente='True')
                    return render(request, "adm_becas/deletipobeca.html", data)
                except Exception as ex:
                    pass

            elif action == 'listautilizacion':
                try:
                    data['title'] = u'Gestión de Utilización de Beca'
                    search = None
                    ids = None
                    if 's' in request.GET:
                        search = request.GET['s']
                        if search.isdigit():
                            utilizacion = BecaUtilizacion.objects.select_related().filter(pk=search,
                                                                                          status=True).order_by('id')
                        else:
                            utilizacion = BecaUtilizacion.objects.select_related().filter(
                                Q(nombre__icontains=search), status=True).order_by('id')
                    elif 'id' in request.GET:
                        ids = request.GET['id']
                        utilizacion = BecaUtilizacion.objects.filter(id=ids, status=True, vigente='True').order_by('id')
                    else:
                        utilizacion = BecaUtilizacion.objects.filter(status=True).order_by('id')
                    paging = MiPaginador(utilizacion, 10)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['page'] = page
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['listautilizacion'] = page.object_list
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    return render(request, "adm_becas/listautilizacion.html", data)
                except Exception as ex:
                    pass

            elif action == 'addutilizacion':
                try:
                    data['title'] = u'Agregar Utilización'
                    form = BecaUtilizacionForm()
                    data['form'] = form
                    return render(request, "adm_becas/addutilizacion.html", data)
                except Exception as ex:
                    pass

            elif action == 'editutilizacion':
                try:
                    data['title'] = u'Editar Utilización'
                    data['utilizacion'] = utilizacion = BecaUtilizacion.objects.get(pk=int(request.GET['id']),
                                                                                    status=True)
                    form = BecaUtilizacionForm(initial={'nombre': utilizacion.nombre,
                                                        'vigente': utilizacion.vigente})
                    data['form'] = form
                    return render(request, "adm_becas/editutilizacion.html", data)
                except Exception as ex:
                    pass

            elif action == 'deleutilizacion':
                try:
                    data['title'] = u'Eliminar Utilización'
                    data['utilizacion'] = BecaUtilizacion.objects.get(pk=int(request.GET['id']), status=True,
                                                                      vigente='True')
                    return render(request, "adm_becas/deleutilizacion.html", data)
                except Exception as ex:
                    pass

            elif action == 'listarequisito':
                try:
                    data['title'] = u'Gestión de Requisitos'
                    search = None
                    ids = None
                    if 's' in request.GET:
                        search = request.GET['s']
                        if search.isdigit():
                            requisito = BecaRequisitos.objects.select_related().filter(pk=search, status=True).order_by(
                                'id')
                        else:
                            requisito = BecaRequisitos.objects.select_related().filter(
                                Q(nombre__icontains=search), status=True).order_by('id')
                    elif 'id' in request.GET:
                        ids = request.GET['id']
                        requisito = BecaRequisitos.objects.filter(id=ids, status=True, vigente='True').order_by('id')
                    else:
                        requisito = BecaRequisitos.objects.filter(status=True).order_by('id')
                    paging = MiPaginador(requisito, 10)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['page'] = page
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['listarequisito'] = page.object_list
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    return render(request, "adm_becas/listarequisito.html", data)
                except Exception as ex:
                    pass

            elif action == 'addrequisito':
                try:
                    data['title'] = u'Agregar Requisito'
                    form = BecaRequisitosForm()
                    data['form'] = form
                    return render(request, "adm_becas/addrequisito.html", data)
                except Exception as ex:
                    pass

            elif action == 'editrequisito':
                try:
                    data['title'] = u'Editar Requisito'
                    data['requisito'] = requisito = BecaRequisitos.objects.get(pk=int(request.GET['id']), status=True)
                    form = BecaRequisitosForm(initial={'nombre': requisito.nombre,
                                                       'vigente': requisito.vigente,
                                                       'matricula': requisito.matricula,
                                                       'regular': requisito.regular,
                                                       'residencia': requisito.residencia,
                                                       'reprobado': requisito.reprobado,
                                                       'nodeudar': requisito.nodeudar,
                                                       })
                    data['form'] = form
                    return render(request, "adm_becas/editrequisito.html", data)
                except Exception as ex:
                    pass

            elif action == 'delerequisito':
                try:
                    data['title'] = u'Eliminar Requisito'
                    data['requisito'] = BecaRequisitos.objects.get(pk=int(request.GET['id']), status=True,
                                                                   vigente='True')
                    return render(request, "adm_becas/delerequisito.html", data)
                except Exception as ex:
                    pass

            elif action == 'listaquintil':
                try:
                    data['tipobeca'] = tipo = BecaTipo.objects.get(status=True, vigente=True, id=18)
                    configuracion = tipo.becatipoconfiguracion_set.filter(becaperiodo__periodo=periodo).first()
                    totalconf_requisitos = configuracion.requisitosbecas.count() if hasattr(configuracion, 'requisitosbecas') else 0
                    data['title'] = tipo.nombre
                    lista = []
                    search = None
                    ids = None
                    nuevolistado = None
                    materias = None
                    if periodosesion.versionbeca is None or periodosesion.versionbeca == 1:
                        if 's' in request.GET:
                            search = request.GET['s'].strip()
                            if search.isdigit():
                                listaquintil = MatriculaGrupoSocioEconomico.objects.values_list(
                                    'matricula__inscripcion__id', flat=True).filter(
                                    Q(matricula__inscripcion__persona__cedula=search),
                                    matricula__estado_matricula__in=[2, 3], matricula__retiradomatricula=False,
                                    matricula__nivel__periodo__id=periodo.id,
                                    matricula__inscripcion__carrera__coordinacion__excluir=False).distinct().order_by(
                                    "puntajetotal")
                        else:
                            listaquintil = MatriculaGrupoSocioEconomico.objects.values_list(
                                'matricula__inscripcion__id', flat=True).filter(
                                Q(gruposocioeconomico__codigo='D') | Q(gruposocioeconomico__codigo='C-'),
                                matricula__estado_matricula__in=[2, 3], matricula__retiradomatricula=False,
                                matricula__nivel__periodo__id=periodo.id,
                                matricula__inscripcion__carrera__coordinacion__excluir=False).distinct().order_by(
                                "puntajetotal")
                        matriculados = MatriculaGrupoSocioEconomico.objects.filter(
                            Q(gruposocioeconomico__codigo='D') | Q(gruposocioeconomico__codigo='C-'),
                            matricula__retiradomatricula=False, matricula__nivel__periodo__id=anterior.id,
                            matricula__matriculagruposocioeconomico__tipomatricula=1,
                            matricula__inscripcion__carrera__coordinacion__excluir=False,
                            matricula__inscripcion__id__in=listaquintil).distinct().order_by(
                            "gruposocioeconomico__codigo")
                    elif periodosesion.versionbeca == 2:
                        inscripciones = PreInscripcionBeca.objects.filter(periodo=periodosesion,
                                                                          becatipo_id=18).values_list('inscripcion_id',
                                                                                                      flat=True)
                        matriculados = MatriculaGrupoSocioEconomico.objects.filter(
                            Q(gruposocioeconomico__codigo='D') | Q(gruposocioeconomico__codigo='C-'),
                            matricula__retiradomatricula=False, matricula__nivel__periodo=periodosesion,
                            matricula__matriculagruposocioeconomico__tipomatricula=1,
                            matricula__inscripcion__in=inscripciones,
                            matricula__inscripcion__carrera__coordinacion__excluir=False).distinct().order_by(
                            "gruposocioeconomico__codigo")
                        if 's' in request.GET:
                            search = request.GET['s'].strip()
                            ss = search.split(' ')
                            if len(ss) == 1:
                                matriculados = matriculados.filter(
                                    Q(matricula__inscripcion__persona__cedula__icontains=search) |
                                    Q(matricula__inscripcion__persona__pasaporte__icontains=search) |
                                    Q(matricula__inscripcion__persona__apellido1__icontains=search) |
                                    Q(matricula__inscripcion__persona__apellido2__icontains=search))
                            else:
                                matriculados = matriculados.filter(
                                    Q(matricula__inscripcion__persona__apellido1__icontains=ss[0]) &
                                    Q(matricula__inscripcion__persona__apellido2__icontains=ss[1]))
                    g_id = 0
                    em_id = 0
                    if 'g' in request.GET:
                        gruposocio = None
                        if request.GET['g']:
                            gruposocio = GrupoSocioEconomico.objects.filter(pk=int(request.GET['g']))
                            if gruposocio.exists():
                                g_id = int(request.GET['g'])
                                matriculados = matriculados.filter(
                                    gruposocioeconomico_id__in=gruposocio.values_list('id', flat=True))
                    if 'em' in request.GET and int(request.GET['em']) > 0:
                        em_id = int(request.GET['em'])
                        matriculados = matriculados.filter(matricula__estado_matricula=em_id)
                    if 'id' in request.GET:
                        ids = request.GET['id']
                        matriculados = matriculados.filter(pk=ids)

                    data['carreras'] = carreras = Carrera.objects.filter(status=True, pk__in=matriculados.values_list(
                        'matricula__inscripcion__carrera_id', flat=True).distinct())
                    data['niveles'] = niveles = NivelMalla.objects.filter(status=True).order_by('orden')

                    resumen = {}
                    resumen['carreras'] = []
                    idmats = matriculados.values_list('matricula__inscripcion_id', flat=True).distinct()
                    estados_matriculas = matriculados.values_list('matricula__estado_matricula', flat=True).order_by('matricula__estado_matricula').distinct()
                    resumen['estados_matriculas'] = [{'name': dict(ESTADO_MATRICULA)[estadomatr], 'count':matriculados.filter(matricula__estado_matricula=estadomatr).count()} for estadomatr in estados_matriculas]
                    resumen['total'] = matriculados.count()
                    resumen['count_renovacion'] = BecaSolicitud.objects.values("id").filter(inscripcion__id__in=idmats, periodo=anterior, becaaceptada=2, becaasignada=2, becatipo=tipo).count()
                    resumen['count_nueva'] = resumen['total'] - resumen['count_renovacion']
                    resumen['totalgeneralrequisitos'] = PreInscripcionBeca.objects.filter(inscripcion_id__in=idmats, periodo=periodo, becatipo=tipo).count()
                    resumen['totalcumplenrequisitos'] = PreInscripcionBeca.objects.filter(inscripcion_id__in=idmats, periodo=periodo, becatipo=tipo, preinscripcionbecarequisito__cumplerequisito=True).annotate(total=Count('preinscripcionbecarequisito__id')).filter(total=totalconf_requisitos).distinct().count()
                    resumen['totalpendientesrequisitos'] = PreInscripcionBeca.objects.filter(inscripcion_id__in=idmats, periodo=periodo, becatipo=tipo, preinscripcionbecarequisito__cumplerequisito__isnull=True).distinct().count()
                    resumen['totalrechazadosrequisitos'] = resumen['totalgeneralrequisitos'] - (resumen['totalcumplenrequisitos'] + resumen['totalpendientesrequisitos'])

                    for carrera in carreras:
                        datacarr = {
                            'name': carrera.nombre,
                            'niveles': []
                        }
                        for nivel in niveles:
                            datacarr['niveles'].append(matriculados.filter(matricula__inscripcion__carrera=carrera, matricula__nivelmalla=nivel).count())
                        datacarr['total'] = sum(datacarr['niveles'])
                        resumen['carreras'].append(datacarr)
                    data['resumen'] = resumen
                    paging = MiPaginador(matriculados, 25)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['page'] = page
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['economica'] = page.object_list
                    data['search'] = search if search else ""
                    data['total'] = matriculados.count()
                    data['ids'] = ids if ids else ""
                    data['g_id'] = g_id
                    data['em_id'] = em_id
                    return render(request, "adm_becas/listaquintil.html", data)
                except Exception as ex:
                    pass

            elif action == 'listapromedio':
                try:
                    data['tipobeca'] = tipo = BecaTipo.objects.get(status=True, vigente=True, id=17)
                    configuracion = tipo.becatipoconfiguracion_set.filter(becaperiodo__periodo=periodo).first()
                    totalconf_requisitos = configuracion.requisitosbecas.count() if hasattr(configuracion, 'requisitosbecas') else 0
                    data['title'] = tipo.nombre
                    search = None
                    ids = None
                    if periodosesion.versionbeca is None or periodosesion.versionbeca == 1:
                        if 's' in request.GET:
                            search = request.GET['s'].strip()
                            inscripcionesactuales = Matricula.objects.values_list('inscripcion__id', flat=True).filter(
                                status=True,
                                inscripcion__persona__cedula=search,
                                nivel__periodo__id=periodo.id,
                                estado_matricula__in=[2, 3],
                                retiradomatricula=False,
                                matriculagruposocioeconomico__tipomatricula=1).exclude(
                                inscripcion__carrera__modalidad=3).distinct().order_by("inscripcion__persona")
                        else:
                            inscripcionesactuales = Matricula.objects.values_list('inscripcion__id', flat=True).filter(
                                status=True,
                                nivel__periodo__id=periodo.id,
                                estado_matricula__in=[2, 3],
                                retiradomatricula=False,
                                matriculagruposocioeconomico__tipomatricula=1).exclude(
                                inscripcion__carrera__modalidad=3).distinct().order_by("inscripcion__persona")
                        matriculados = Matricula.objects.filter(status=True,
                                                                nivel__periodo__id=periodo.id,
                                                                estado_matricula__in=[2, 3],
                                                                retiradomatricula=False,
                                                                matriculagruposocioeconomico__tipomatricula=1,
                                                                inscripcion__id__in=inscripcionesactuales).exclude(
                            inscripcion__carrera__modalidad=3).distinct().order_by("inscripcion__persona")
                    elif periodosesion.versionbeca == 2:
                        inscripciones = PreInscripcionBeca.objects.filter(periodo=periodosesion,
                                                                          becatipo_id=17).values_list('inscripcion_id',
                                                                                                      flat=True)
                        matriculados = Matricula.objects.filter(inscripcion__id__in=inscripciones,
                                                                nivel__periodo__id=periodo.id).distinct().order_by(
                            "inscripcion__persona")
                        if 's' in request.GET:
                            search = request.GET['s'].strip()
                            ss = search.split(' ')
                            if len(ss) == 1:
                                matriculados = matriculados.filter(Q(inscripcion__persona__cedula__icontains=search) |
                                                                   Q(inscripcion__persona__pasaporte__icontains=search) |
                                                                   Q(inscripcion__persona__apellido1__icontains=search) |
                                                                   Q(inscripcion__persona__apellido2__icontains=search))
                            else:
                                matriculados = matriculados.filter(Q(inscripcion__persona__apellido1__icontains=ss[0]) &
                                                                   Q(inscripcion__persona__apellido2__icontains=ss[1]))
                    if 'id' in request.GET:
                        ids = request.GET['id']
                        matriculados = matriculados.filter(pk=ids)
                    em_id = 0
                    if 'em' in request.GET and int(request.GET['em']) > 0:
                        em_id = int(request.GET['em'])
                        matriculados = matriculados.filter(estado_matricula=em_id)
                    data['carreras'] = carreras = Carrera.objects.filter(status=True, pk__in=matriculados.values_list(
                        'inscripcion__carrera_id', flat=True).distinct())
                    data['niveles'] = niveles = NivelMalla.objects.filter(status=True).order_by('orden')

                    resumen = {}
                    resumen['carreras'] = []
                    idmats = matriculados.values_list('inscripcion_id', flat=True).distinct()
                    estados_matriculas = matriculados.values_list('estado_matricula', flat=True).order_by('estado_matricula').distinct()
                    resumen['estados_matriculas'] = [{'name': dict(ESTADO_MATRICULA)[estadomatr], 'count':matriculados.filter(estado_matricula=estadomatr).count()} for estadomatr in estados_matriculas]
                    resumen['total'] = matriculados.count()
                    resumen['count_renovacion'] = BecaSolicitud.objects.values("id").filter(inscripcion__id__in=idmats, periodo=anterior, becaaceptada=2, becaasignada=2, becatipo=tipo).count()
                    resumen['count_nueva'] = resumen['total'] - resumen['count_renovacion']
                    resumen['totalgeneralrequisitos'] = PreInscripcionBeca.objects.filter(inscripcion_id__in=idmats, periodo=periodo, becatipo=tipo).count()
                    resumen['totalcumplenrequisitos'] = PreInscripcionBeca.objects.filter(inscripcion_id__in=idmats, periodo=periodo, becatipo=tipo, preinscripcionbecarequisito__cumplerequisito=True).annotate(total=Count('preinscripcionbecarequisito__id')).filter(total=totalconf_requisitos).distinct().count()
                    resumen['totalpendientesrequisitos'] = PreInscripcionBeca.objects.filter(inscripcion_id__in=idmats, periodo=periodo, becatipo=tipo, preinscripcionbecarequisito__cumplerequisito__isnull=True).distinct().count()
                    resumen['totalrechazadosrequisitos'] = resumen['totalgeneralrequisitos'] - (resumen['totalcumplenrequisitos'] + resumen['totalpendientesrequisitos'])

                    for carrera in carreras:
                        datacarr = {
                            'name': carrera.nombre,
                            'niveles': []
                        }
                        for nivel in niveles:
                            datacarr['niveles'].append(matriculados.filter(inscripcion__carrera=carrera, nivelmalla=nivel).count())
                        datacarr['total'] = sum(datacarr['niveles'])
                        resumen['carreras'].append(datacarr)
                    data['resumen'] = resumen

                    paging = MiPaginador(matriculados, 25)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    data['paginador'] = p
                    data['paging'] = paging
                    data['page'] = page
                    # data['discapacitados'] = nuevolistado
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['promedioalto'] = page.object_list
                    # data['promedioalto'] = nuevolistado
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['em_id'] = em_id
                    data['total'] = matriculados.count()
                    return render(request, "adm_becas/listapromedio.html", data)
                except Exception as ex:
                    pass

            elif action == 'listadeporte':
                try:
                    data['tipobeca'] = tipo = BecaTipo.objects.get(status=True, vigente=True, id=20)
                    configuracion = tipo.becatipoconfiguracion_set.filter(becaperiodo__periodo=periodo).first()
                    totalconf_requisitos = configuracion.requisitosbecas.count() if hasattr(configuracion, 'requisitosbecas') else 0
                    data['title'] = tipo.nombre
                    search = None
                    ids = None
                    if periodosesion.versionbeca is None or periodosesion.versionbeca == 1:
                        matriculados = Matricula.objects.filter(pk=None)
                    elif periodosesion.versionbeca == 2:
                        inscripciones = PreInscripcionBeca.objects.filter(periodo=periodosesion,
                                                                          becatipo_id=20).values_list('inscripcion_id',
                                                                                                      flat=True)
                        matriculados = Matricula.objects.filter(inscripcion__id__in=inscripciones,
                                                                nivel__periodo__id=periodo.id).distinct().order_by(
                            "inscripcion__persona")
                        if 's' in request.GET:
                            search = request.GET['s'].strip()
                            ss = search.split(' ')
                            if len(ss) == 1:
                                matriculados = matriculados.filter(Q(inscripcion__persona__cedula__icontains=search) |
                                                                   Q(inscripcion__persona__pasaporte__icontains=search) |
                                                                   Q(inscripcion__persona__apellido1__icontains=search) |
                                                                   Q(inscripcion__persona__apellido2__icontains=search))
                            else:
                                matriculados = matriculados.filter(Q(inscripcion__persona__apellido1__icontains=ss[0]) &
                                                                   Q(inscripcion__persona__apellido2__icontains=ss[1]))
                    if 'id' in request.GET:
                        ids = request.GET['id']
                        matriculados = matriculados.filter(pk=ids)

                    data['carreras'] = carreras = Carrera.objects.filter(status=True, pk__in=matriculados.values_list(
                        'inscripcion__carrera_id', flat=True).distinct())
                    data['niveles'] = niveles = NivelMalla.objects.filter(status=True).order_by('orden')

                    resumen = {}
                    resumen['carreras'] = []
                    idmats = matriculados.values_list('inscripcion_id', flat=True).distinct()
                    estados_matriculas = matriculados.values_list('estado_matricula', flat=True).order_by('estado_matricula').distinct()
                    resumen['estados_matriculas'] = [{'name': dict(ESTADO_MATRICULA)[estadomatr], 'count':matriculados.filter(estado_matricula=estadomatr).count()} for estadomatr in estados_matriculas]
                    resumen['total'] = matriculados.count()
                    resumen['count_renovacion'] = BecaSolicitud.objects.values("id").filter(inscripcion__id__in=idmats, periodo=anterior, becaaceptada=2, becaasignada=2, becatipo=tipo).count()
                    resumen['count_nueva'] = resumen['total'] - resumen['count_renovacion']
                    resumen['totalgeneralrequisitos'] = PreInscripcionBeca.objects.filter(inscripcion_id__in=idmats, periodo=periodo, becatipo=tipo).count()
                    resumen['totalcumplenrequisitos'] = PreInscripcionBeca.objects.filter(inscripcion_id__in=idmats, periodo=periodo, becatipo=tipo, preinscripcionbecarequisito__cumplerequisito=True).annotate(total=Count('preinscripcionbecarequisito__id')).filter(total=totalconf_requisitos).distinct().count()
                    resumen['totalpendientesrequisitos'] = PreInscripcionBeca.objects.filter(inscripcion_id__in=idmats, periodo=periodo, becatipo=tipo, preinscripcionbecarequisito__cumplerequisito__isnull=True).distinct().count()
                    resumen['totalrechazadosrequisitos'] = resumen['totalgeneralrequisitos'] - (resumen['totalcumplenrequisitos'] + resumen['totalpendientesrequisitos'])

                    for carrera in carreras:
                        datacarr = {
                            'name': carrera.nombre,
                            'niveles': []
                        }
                        for nivel in niveles:
                            datacarr['niveles'].append(matriculados.filter(inscripcion__carrera=carrera, nivelmalla=nivel).count())
                        datacarr['total'] = sum(datacarr['niveles'])
                        resumen['carreras'].append(datacarr)
                    data['resumen'] = resumen

                    paging = MiPaginador(matriculados, 25)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['page'] = page
                    # data['discapacitados'] = nuevolistado
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['deportes'] = page.object_list
                    # data['promedioalto'] = nuevolistado
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['total'] = matriculados.count()
                    return render(request, "adm_becas/listadeporte.html", data)
                except Exception as ex:
                    pass

            elif action == 'listadiscapacitados':
                try:
                    data['tipobeca'] = tipo = BecaTipo.objects.get(status=True, vigente=True, id=19)
                    configuracion = tipo.becatipoconfiguracion_set.filter(becaperiodo__periodo=periodo).first()
                    totalconf_requisitos = configuracion.requisitosbecas.count() if hasattr(configuracion, 'requisitosbecas') else 0
                    data['title'] = tipo.nombre
                    search = None
                    ids = None
                    if periodosesion.versionbeca is None or periodosesion.versionbeca == 1:
                        if 's' in request.GET:
                            search = request.GET['s'].strip()
                            if search.isdigit():
                                discapacitados = Matricula.objects.values_list('inscripcion__id', flat=True).filter(
                                    status=True, inscripcion__persona__cedula=search,
                                    nivel__periodo__id=periodo.id,
                                    estado_matricula__in=[2, 3],
                                    retiradomatricula=False,
                                    inscripcion__persona__perfilinscripcion__tienediscapacidad=True,
                                    inscripcion__persona__perfilinscripcion__verificadiscapacidad=True,
                                    matriculagruposocioeconomico__tipomatricula=1).order_by("inscripcion__persona")
                        else:
                            discapacitados = Matricula.objects.values_list('inscripcion__id', flat=True).filter(
                                status=True,
                                nivel__periodo__id=periodo.id,
                                estado_matricula__in=[2, 3],
                                retiradomatricula=False,
                                inscripcion__persona__perfilinscripcion__tienediscapacidad=True,
                                inscripcion__persona__perfilinscripcion__verificadiscapacidad=True,
                                matriculagruposocioeconomico__tipomatricula=1).order_by("inscripcion__persona")
                        matriculados = Matricula.objects.filter(status=True, nivel__periodo__id=anterior.id,
                                                                estado_matricula__in=[2, 3],
                                                                retiradomatricula=False,
                                                                inscripcion__persona__perfilinscripcion__tienediscapacidad=True,
                                                                inscripcion__persona__perfilinscripcion__verificadiscapacidad=True,
                                                                matriculagruposocioeconomico__tipomatricula=1,
                                                                inscripcion__id__in=discapacitados).order_by(
                            "inscripcion__persona")
                    elif periodosesion.versionbeca == 2:
                        inscripciones = PreInscripcionBeca.objects.filter(periodo=periodosesion,becatipo_id=19).values_list('inscripcion_id', flat=True)
                        matriculados = Matricula.objects.filter(inscripcion__id__in=inscripciones,
                                                                nivel__periodo__id=periodo.id).distinct().order_by("inscripcion__persona")
                        if 's' in request.GET:
                            search = request.GET['s'].strip()
                            ss = search.split(' ')
                            if len(ss) == 1:
                                matriculados = matriculados.filter(Q(inscripcion__persona__cedula__icontains=search) |
                                                                   Q(inscripcion__persona__pasaporte__icontains=search) |
                                                                   Q(inscripcion__persona__apellido1__icontains=search) |
                                                                   Q(inscripcion__persona__apellido2__icontains=search))
                            else:
                                matriculados = matriculados.filter(Q(inscripcion__persona__apellido1__icontains=ss[0]) &
                                                                   Q(inscripcion__persona__apellido2__icontains=ss[1]))

                    preinscripciones = PreInscripcionBeca.objects.filter(periodo=periodosesion, becatipo_id=19, inscripcion_id__in=matriculados.values_list('inscripcion_id', flat=True))
                    d_id = 0
                    em_id = 0
                    data['discapacidades'] = discapacidades = Discapacidad.objects.filter(status=True)
                    if 'd' in request.GET and int(request.GET['d']) > 0:
                        if discapacidades.filter(pk=int(request.GET['d'])).exists():
                            d_id = int(request.GET['d'])
                            if periodo >= 119:
                                matriculados = matriculados.filter(inscripcion_id__in=preinscripciones.filter(tipodiscapacidad_id=d_id).values_list('inscripcion_id', flat=True))
                            else:
                                matriculados = matriculados.filter(inscripcion__persona__perfilinscripcion__tipodiscapacidad=discapacidades.filter(pk=int(request.GET['d'])).first())
                    if 'em' in request.GET and int(request.GET['em']) > 0:
                        em_id = int(request.GET['em'])
                        matriculados = matriculados.filter(estado_matricula=em_id)
                    if 'id' in request.GET:
                        ids = request.GET['id']
                        matriculados = matriculados.filter(pk=ids)

                    data['carreras'] = carreras = Carrera.objects.filter(status=True, pk__in=matriculados.values_list('inscripcion__carrera_id', flat=True).distinct())
                    data['niveles'] = niveles = NivelMalla.objects.filter(status=True).order_by('orden')

                    resumen = {}
                    resumen['carreras'] = []
                    idmats = matriculados.values_list('inscripcion_id', flat=True).distinct()
                    estados_matriculas = matriculados.values_list('estado_matricula', flat=True).order_by('estado_matricula').distinct()
                    resumen['estados_matriculas'] = [{'name': dict(ESTADO_MATRICULA)[estadomatr], 'count':matriculados.filter(estado_matricula=estadomatr).count()} for estadomatr in estados_matriculas]
                    resumen['total'] = matriculados.count()
                    resumen['count_renovacion'] = BecaSolicitud.objects.values("id").filter(inscripcion__id__in=matriculados.values_list('inscripcion_id', flat=True).distinct(), periodo=anterior, becaaceptada=2, becaasignada=2, becatipo=tipo).count()
                    resumen['count_nueva'] = resumen['total'] - resumen['count_renovacion']
                    resumen['totalgeneralrequisitos'] = PreInscripcionBeca.objects.filter(inscripcion_id__in=idmats, periodo=periodo, becatipo=tipo).count()
                    resumen['totalcumplenrequisitos'] = PreInscripcionBeca.objects.filter(inscripcion_id__in=idmats, periodo=periodo, becatipo=tipo, preinscripcionbecarequisito__cumplerequisito=True).annotate(total=Count('preinscripcionbecarequisito__id')).filter(total=totalconf_requisitos).distinct().count()
                    resumen['totalpendientesrequisitos'] = PreInscripcionBeca.objects.filter(inscripcion_id__in=idmats, periodo=periodo, becatipo=tipo, preinscripcionbecarequisito__cumplerequisito__isnull=True).distinct().count()
                    resumen['totalrechazadosrequisitos'] = resumen['totalgeneralrequisitos'] - (resumen['totalcumplenrequisitos'] + resumen['totalpendientesrequisitos'])

                    for carrera in carreras:
                        datacarr = {
                            'name': carrera.nombre,
                            'niveles': []
                        }
                        for nivel in niveles:
                            datacarr['niveles'].append(matriculados.filter(inscripcion__carrera=carrera, nivelmalla=nivel).count())
                        datacarr['total'] = sum(datacarr['niveles'])
                        resumen['carreras'].append(datacarr)
                    data['resumen'] = resumen

                    paging = MiPaginador(matriculados, 25)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['page'] = page
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['discapacitados'] = page.object_list
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['d_id'] = d_id
                    data['em_id'] = em_id
                    data['total'] = matriculados.count()
                    return render(request, "adm_becas/listadiscapacitados.html", data)
                except Exception as ex:
                    pass

            elif action == 'listaetnia':
                try:
                    data['tipobeca'] = tipo = BecaTipo.objects.get(status=True, vigente=True, id=21)
                    configuracion = tipo.becatipoconfiguracion_set.filter(becaperiodo__periodo=periodo).first()
                    totalconf_requisitos = configuracion.requisitosbecas.count() if hasattr(configuracion, 'requisitosbecas') else 0
                    data['title'] = tipo.nombre
                    search = None
                    ids = None
                    if periodosesion.versionbeca is None or periodosesion.versionbeca == 1:
                        if 's' in request.GET:
                            search = request.GET['s'].strip()
                            if search.isdigit():
                                etnias = Matricula.objects.values_list('inscripcion__id', flat=True).filter(
                                    status=True, inscripcion__persona__cedula=search,
                                    nivel__periodo__id=periodo.id,
                                    estado_matricula__in=[2, 3],
                                    retiradomatricula=False,
                                    inscripcion__persona__perfilinscripcion__nacionalidadindigena__id__in=[1, 14],
                                    matriculagruposocioeconomico__tipomatricula=1).order_by("inscripcion__persona")
                        else:
                            etnias = Matricula.objects.values_list('inscripcion__id', flat=True).filter(status=True,
                                                                                                        nivel__periodo__id=periodo.id,
                                                                                                        estado_matricula__in=[2, 3],
                                                                                                        retiradomatricula=False,
                                                                                                        matriculagruposocioeconomico__tipomatricula=1,
                                                                                                        inscripcion__persona__perfilinscripcion__nacionalidadindigena__id=7).order_by("inscripcion__persona")
                        matriculados = Matricula.objects.filter(status=True, nivel__periodo__id=anterior.id,
                                                                estado_matricula__in=[2, 3],
                                                                retiradomatricula=False,
                                                                inscripcion__id__in=etnias,
                                                                matriculagruposocioeconomico__tipomatricula=1).order_by("inscripcion__persona")
                    elif periodosesion.versionbeca == 2:
                        inscripciones = PreInscripcionBeca.objects.filter(periodo=periodosesion, becatipo_id=21)
                        preinscripciones = inscripciones
                        if periodo.id >= 119:
                            data['razas'] = razas = Raza.objects.filter(status=True, pk__in=inscripciones.values_list('raza_id', flat=True).distinct())
                            if 'r' in request.GET and int(request.GET['r']) > 0:
                                raza = razas.filter(pk=int(request.GET['r'])).first()
                                if raza is not None:
                                    inscripciones = inscripciones.filter(raza=raza)
                        inscripciones = inscripciones.values_list('inscripcion_id',flat=True)
                        matriculados = Matricula.objects.filter(inscripcion__id__in=inscripciones,
                                                                nivel__periodo__id=periodo.id).distinct().order_by("inscripcion__persona")
                        if 's' in request.GET:
                            search = request.GET['s'].strip()
                            ss = search.split(' ')
                            if len(ss) == 1:
                                matriculados = matriculados.filter(Q(inscripcion__persona__cedula__icontains=search) |
                                                                   Q(inscripcion__persona__pasaporte__icontains=search) |
                                                                   Q(inscripcion__persona__apellido1__icontains=search) |
                                                                   Q(inscripcion__persona__apellido2__icontains=search))
                            else:
                                matriculados = matriculados.filter(Q(inscripcion__persona__apellido1__icontains=ss[0]) &
                                                                   Q(inscripcion__persona__apellido2__icontains=ss[1]))
                    r_id = 0
                    em_id = 0
                    if periodo.id < 119:
                        data['razas'] = razas = Raza.objects.filter(status=True, pk__in=inscripciones.values_list('inscripcion__persona__perfilinscripcion__raza_id', flat=True).distinct())
                        if 'r' in request.GET and int(request.GET['r']) > 0:
                            if razas.filter(pk=int(request.GET['r'])).exists():
                                r_id = int(request.GET['r'])
                                matriculados = matriculados.filter(inscripcion__persona__perfilinscripcion__raza=razas.filter(pk=int(request.GET['r'])).first())

                    if 'em' in request.GET and int(request.GET['em']) > 0:
                        em_id = int(request.GET['em'])
                        matriculados = matriculados.filter(estado_matricula=em_id)
                    if 'id' in request.GET:
                        ids = request.GET['id']
                        matriculados = matriculados.filter(pk=ids)

                    data['carreras'] = carreras = Carrera.objects.filter(status=True, pk__in=matriculados.values_list(
                        'inscripcion__carrera_id', flat=True).distinct())
                    data['niveles'] = niveles = NivelMalla.objects.filter(status=True).order_by('orden')

                    resumen = {}
                    resumen['carreras'] = []
                    idmats = matriculados.values_list('inscripcion_id', flat=True).distinct()
                    estados_matriculas = matriculados.values_list('estado_matricula', flat=True).order_by('estado_matricula').distinct()
                    resumen['estados_matriculas'] = [{'name': dict(ESTADO_MATRICULA)[estadomatr], 'count':matriculados.filter(estado_matricula=estadomatr).count()} for estadomatr in estados_matriculas]
                    if periodo.id < 119:
                        resumen['razas'] = [{'name': raza.__str__(), 'count': matriculados.filter(inscripcion__persona__perfilinscripcion__raza=raza).count()} for raza in razas]
                    else:
                        resumen['razas'] = [{'name': raza.__str__(), 'count': preinscripciones.filter(raza=raza).count()} for raza in razas]
                    resumen['total'] = matriculados.count()
                    resumen['count_renovacion'] = BecaSolicitud.objects.values("id").filter(inscripcion__id__in=matriculados.values_list('inscripcion_id', flat=True).distinct(), periodo=anterior, becaaceptada=2, becaasignada=2, becatipo=tipo).count()
                    resumen['count_nueva'] = resumen['total'] - resumen['count_renovacion']
                    resumen['totalgeneralrequisitos'] = PreInscripcionBeca.objects.filter(inscripcion_id__in=idmats, periodo=periodo, becatipo=tipo).count()
                    resumen['totalcumplenrequisitos'] = PreInscripcionBeca.objects.filter(inscripcion_id__in=idmats, periodo=periodo, becatipo=tipo, preinscripcionbecarequisito__cumplerequisito=True).annotate(total=Count('preinscripcionbecarequisito__id')).filter(total=totalconf_requisitos).distinct().count()
                    resumen['totalpendientesrequisitos'] = PreInscripcionBeca.objects.filter(inscripcion_id__in=idmats, periodo=periodo, becatipo=tipo, preinscripcionbecarequisito__cumplerequisito__isnull=True).distinct().count()
                    resumen['totalrechazadosrequisitos'] = resumen['totalgeneralrequisitos'] - (resumen['totalcumplenrequisitos'] + resumen['totalpendientesrequisitos'])

                    for carrera in carreras:
                        datacarr = {
                            'name': carrera.nombre,
                            'niveles': []
                        }
                        for nivel in niveles:
                            datacarr['niveles'].append(matriculados.filter(inscripcion__carrera=carrera, nivelmalla=nivel).count())
                        datacarr['total'] = sum(datacarr['niveles'])
                        resumen['carreras'].append(datacarr)
                    data['resumen'] = resumen

                    paging = MiPaginador(matriculados, 25)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['page'] = page
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['etnias'] = page.object_list
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['r_id'] = r_id
                    data['em_id'] = em_id
                    data['total'] = matriculados.count()
                    return render(request, "adm_becas/listaetnias.html", data)
                except Exception as ex:
                    pass

            elif action == 'listaextranjero':
                try:
                    data['tipobeca'] = tipo = BecaTipo.objects.get(status=True, vigente=True, id=22)
                    configuracion = tipo.becatipoconfiguracion_set.filter(becaperiodo__periodo=periodo).first()
                    totalconf_requisitos = configuracion.requisitosbecas.count() if hasattr(configuracion, 'requisitosbecas') else 0
                    data['title'] = tipo.nombre
                    search = None
                    ids = None
                    if periodosesion.versionbeca is None or periodosesion.versionbeca == 1:
                        matriculados = Matricula.objects.filter(pk=None)
                    elif periodosesion.versionbeca == 2:
                        inscripciones = PreInscripcionBeca.objects.filter(periodo=periodosesion,
                                                                          becatipo_id=22).values_list('inscripcion_id',
                                                                                                      flat=True)
                        matriculados = Matricula.objects.filter(~Q(inscripcion__persona__pais_id=1),
                                                                inscripcion__id__in=inscripciones,
                                                                nivel__periodo__id=periodo.id).distinct().order_by("inscripcion__persona")
                        if 's' in request.GET:
                            search = request.GET['s'].strip()
                            ss = search.split(' ')
                            if len(ss) == 1:
                                matriculados = matriculados.filter(Q(inscripcion__persona__cedula__icontains=search) |
                                                                   Q(inscripcion__persona__pasaporte__icontains=search) |
                                                                   Q(inscripcion__persona__apellido1__icontains=search) |
                                                                   Q(inscripcion__persona__apellido2__icontains=search))
                            else:
                                matriculados = matriculados.filter(Q(inscripcion__persona__apellido1__icontains=ss[0]) &
                                                                   Q(inscripcion__persona__apellido2__icontains=ss[1]))
                    p_id = 0
                    em_id = 0

                    data['paises'] = paises = Pais.objects.filter(status=True, pk__in=inscripciones.values_list(
                        'inscripcion__persona__pais_id', flat=True).distinct()).exclude(id=1)
                    if 'p' in request.GET and int(request.GET['p']) > 0:
                        if paises.filter(pk=int(request.GET['p'])).exists():
                            r_id = int(request.GET['p'])
                            matriculados = matriculados.filter(
                                inscripcion__persona__pais=paises.filter(pk=int(request.GET['p'])).first())
                    if 'em' in request.GET and int(request.GET['em']) > 0:
                        em_id = int(request.GET['em'])
                        matriculados = matriculados.filter(estado_matricula=em_id)
                    if 'id' in request.GET:
                        ids = request.GET['id']
                        matriculados = matriculados.filter(pk=ids)

                    paisesfiltros = Pais.objects.filter(status=True, pk__in=inscripciones.values_list('inscripcion__persona__pais_id', flat=True).distinct()).exclude(id=1)
                    data['carreras'] = carreras = Carrera.objects.filter(status=True, pk__in=matriculados.values_list('inscripcion__carrera_id', flat=True).distinct())
                    data['niveles'] = niveles = NivelMalla.objects.filter(status=True).order_by('orden')

                    resumen = {}
                    resumen['carreras'] = []
                    idmats = matriculados.values_list('inscripcion_id', flat=True).distinct()
                    estados_matriculas = matriculados.values_list('estado_matricula', flat=True).order_by('estado_matricula').distinct()
                    resumen['estados_matriculas'] = [{'name': dict(ESTADO_MATRICULA)[estadomatr], 'count':matriculados.filter(estado_matricula=estadomatr).count()} for estadomatr in estados_matriculas]
                    resumen['paises'] = [{'name': pais.__str__(), 'count': matriculados.filter(inscripcion__persona__pais=pais).count()} for pais in paisesfiltros]

                    resumen['total'] = matriculados.count()
                    resumen['count_renovacion'] = BecaSolicitud.objects.values("id").filter(inscripcion__id__in=matriculados.values_list('inscripcion_id', flat=True).distinct(), periodo=anterior, becaaceptada=2, becaasignada=2, becatipo=tipo).count()
                    resumen['count_nueva'] = resumen['total'] - resumen['count_renovacion']
                    resumen['totalgeneralrequisitos'] = PreInscripcionBeca.objects.filter(inscripcion_id__in=idmats, periodo=periodo, becatipo=tipo).count()
                    resumen['totalcumplenrequisitos'] = PreInscripcionBeca.objects.filter(inscripcion_id__in=idmats, periodo=periodo, becatipo=tipo, preinscripcionbecarequisito__cumplerequisito=True).annotate(total=Count('preinscripcionbecarequisito__id')).filter(total=totalconf_requisitos).distinct().count()
                    resumen['totalpendientesrequisitos'] = PreInscripcionBeca.objects.filter(inscripcion_id__in=idmats, periodo=periodo, becatipo=tipo, preinscripcionbecarequisito__cumplerequisito__isnull=True).distinct().count()
                    resumen['totalrechazadosrequisitos'] = resumen['totalgeneralrequisitos'] - (resumen['totalcumplenrequisitos'] + resumen['totalpendientesrequisitos'])

                    for carrera in carreras:
                        datacarr = {
                            'name': carrera.nombre,
                            'niveles': []
                        }
                        for nivel in niveles:
                            datacarr['niveles'].append(matriculados.filter(inscripcion__carrera=carrera, nivelmalla=nivel).count())
                        datacarr['total'] = sum(datacarr['niveles'])
                        resumen['carreras'].append(datacarr)
                    data['resumen'] = resumen


                    paging = MiPaginador(matriculados, 25)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['page'] = page
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['matriculados'] = page.object_list
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['p_id'] = p_id
                    data['em_id'] = em_id
                    data['total'] = matriculados.count()
                    return render(request, "adm_becas/listaextranjeros.html", data)
                except Exception as ex:
                    pass

            elif action == 'listamigrante':
                try:
                    data['tipobeca'] = tipo = BecaTipo.objects.get(status=True, vigente=True, id=22)
                    configuracion = tipo.becatipoconfiguracion_set.filter(becaperiodo__periodo=periodo).first()
                    totalconf_requisitos = configuracion.requisitosbecas.count() if hasattr(configuracion, 'requisitosbecas') else 0
                    data['title'] = tipo.nombre
                    search = None
                    ids = None
                    if periodosesion.versionbeca is None or periodosesion.versionbeca == 1:
                        matriculados = Matricula.objects.filter(pk=None)
                    elif periodosesion.versionbeca == 2:
                        inscripciones = PreInscripcionBeca.objects.filter(periodo=periodosesion,
                                                                          becatipo_id=22).values_list('inscripcion_id',
                                                                                                      flat=True)
                        matriculados = Matricula.objects.filter(inscripcion__persona__migrantepersona__isnull=False,
                                                                inscripcion__persona__pais_id=1,
                                                                inscripcion__id__in=inscripciones,
                                                                nivel__periodo__id=periodo.id).distinct().order_by(
                            "inscripcion__persona")
                        if 's' in request.GET:
                            search = request.GET['s'].strip()
                            ss = search.split(' ')
                            if len(ss) == 1:
                                matriculados = matriculados.filter(Q(inscripcion__persona__cedula__icontains=search) |
                                                                   Q(inscripcion__persona__pasaporte__icontains=search) |
                                                                   Q(inscripcion__persona__apellido1__icontains=search) |
                                                                   Q(inscripcion__persona__apellido2__icontains=search))
                            else:
                                matriculados = matriculados.filter(Q(inscripcion__persona__apellido1__icontains=ss[0]) &
                                                                   Q(inscripcion__persona__apellido2__icontains=ss[1]))
                    p_id = 0
                    em_id = 0
                    data['paises'] = paises = Pais.objects.filter(status=True, pk__in=inscripciones.values_list(
                        'inscripcion__persona__migrantepersona__paisresidencia_id', flat=True).distinct())
                    if 'p' in request.GET and int(request.GET['p']) > 0:
                        if paises.filter(pk=int(request.GET['p'])).exists():
                            r_id = int(request.GET['p'])
                            matriculados = matriculados.filter(
                                inscripcion__persona__perfilinscripcion__raza=paises.filter(
                                    pk=int(request.GET['p'])).first())
                    if 'em' in request.GET and int(request.GET['em']) > 0:
                        em_id = int(request.GET['em'])
                        matriculados = matriculados.filter(estado_matricula=em_id)
                    if 'id' in request.GET:
                        ids = request.GET['id']
                        matriculados = matriculados.filter(pk=ids)

                    data['carreras'] = carreras = Carrera.objects.filter(status=True, pk__in=matriculados.values_list(
                        'inscripcion__carrera_id', flat=True).distinct())
                    data['niveles'] = niveles = NivelMalla.objects.filter(status=True).order_by('orden')

                    resumen = {}
                    resumen['carreras'] = []
                    idmats = matriculados.values_list('inscripcion_id', flat=True).distinct()
                    estados_matriculas = matriculados.values_list('estado_matricula', flat=True).order_by('estado_matricula').distinct()
                    resumen['estados_matriculas'] = [{'name': dict(ESTADO_MATRICULA)[estadomatr], 'count':matriculados.filter(estado_matricula=estadomatr).count()} for estadomatr in estados_matriculas]
                    resumen['total'] = matriculados.count()
                    resumen['count_renovacion'] = BecaSolicitud.objects.values("id").filter(inscripcion__id__in=matriculados.values_list('inscripcion_id', flat=True).distinct(), periodo=anterior, becaaceptada=2, becaasignada=2, becatipo=tipo).count()
                    resumen['count_nueva'] = resumen['total'] - resumen['count_renovacion']
                    resumen['totalgeneralrequisitos'] = PreInscripcionBeca.objects.filter(inscripcion_id__in=idmats, periodo=periodo, becatipo=tipo).count()
                    resumen['totalcumplenrequisitos'] = PreInscripcionBeca.objects.filter(inscripcion_id__in=idmats, periodo=periodo, becatipo=tipo, preinscripcionbecarequisito__cumplerequisito=True).annotate(total=Count('preinscripcionbecarequisito__id')).filter(total=totalconf_requisitos).distinct().count()
                    resumen['totalpendientesrequisitos'] = PreInscripcionBeca.objects.filter(inscripcion_id__in=idmats, periodo=periodo, becatipo=tipo, preinscripcionbecarequisito__cumplerequisito__isnull=True).distinct().count()
                    resumen['totalrechazadosrequisitos'] = resumen['totalgeneralrequisitos'] - (resumen['totalcumplenrequisitos'] + resumen['totalpendientesrequisitos'])

                    for carrera in carreras:
                        datacarr = {
                            'name': carrera.nombre,
                            'niveles': []
                        }
                        for nivel in niveles:
                            datacarr['niveles'].append(matriculados.filter(inscripcion__carrera=carrera, nivelmalla=nivel).count())
                        datacarr['total'] = sum(datacarr['niveles'])
                        resumen['carreras'].append(datacarr)
                    data['resumen'] = resumen
                    paging = MiPaginador(matriculados, 25)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['page'] = page
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['matriculados'] = page.object_list
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['p_id'] = p_id
                    data['em_id'] = em_id
                    data['total'] = matriculados.count()
                    return render(request, "adm_becas/listamigrantes.html", data)
                except Exception as ex:
                    pass

            elif action == 'listasolicitudes':
                try:
                    data['title'] = u'SOLICITUDES DE BECAS'
                    search = None
                    ids = None
                    solicitudes = BecaSolicitud.objects.filter(status=True, periodo=periodo,
                                                               periodocalifica__isnull=False).order_by(
                        '-fecha_creacion')
                    if 'id' in request.GET:
                        ids = request.GET['id']
                        solicitudes = solicitudes.filter(id=search)
                    if 's' in request.GET:
                        search = request.GET['s'].strip()
                        if search.isdigit():
                            solicitudes = solicitudes.filter(Q(inscripcion__persona__cedula=search) | Q(id=int(search)))
                        else:
                            ss = search.split(' ')
                            if len(ss) == 1:
                                solicitudes = solicitudes.filter(Q(inscripcion__persona__nombres__icontains=search) | Q(
                                    inscripcion__persona__apellido1__icontains=search) | Q(
                                    inscripcion__persona__apellido2__icontains=search) | Q(
                                    becatipo__nombre__icontains=search))
                            else:
                                solicitudes = solicitudes.filter(
                                    Q(inscripcion__persona__apellido1__icontains=ss[0]) | Q(
                                        becatipo__nombre__icontains=ss[0]) | Q(
                                        inscripcion__persona__apellido2__icontains=ss[1]) | Q(
                                        becatipo__nombre__icontains=ss[1]))
                    btipobeca = 0
                    bestadobeca = 0
                    bestadoaceptado = 0
                    bestadoaasignado = 0
                    if 'btipobeca' in request.GET:
                        btipobeca = int(request.GET['btipobeca'])
                        if btipobeca > 0:
                            solicitudes = solicitudes.filter(becatipo_id=btipobeca)

                    if 'bestadobeca' in request.GET:
                        bestadobeca = int(request.GET['bestadobeca'])
                        if bestadobeca > 0:
                            solicitudes = solicitudes.filter(estado=bestadobeca)

                    if 'bestadoaceptado' in request.GET:
                        bestadoaceptado = int(request.GET['bestadoaceptado'])
                        if bestadoaceptado > 0:
                            solicitudes = solicitudes.filter(becaaceptada=bestadoaceptado)

                    if 'bestadoaasignado' in request.GET:
                        bestadoaasignado = int(request.GET['bestadoaasignado'])
                        if bestadoaasignado > 0:
                            solicitudes = solicitudes.filter(becaasignada=bestadoaasignado)

                    paging = MiPaginador(solicitudes, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['page'] = page
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['solicitudes'] = page.object_list
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['fechaactual'] = datetime.now().strftime('%d-%m-%Y')
                    data['totalsolicitud'] = total = solicitudes.count()
                    data['totaldesistieron'] = totaldesiste = solicitudes.filter(estado=8).count()
                    data['totalneto'] = totalneto = total - totaldesiste
                    data['totalrevision'] = totalrevision = solicitudes.filter(estado=4).count()
                    data['totalaprobada'] = totalaprobada = solicitudes.filter(estado=2).count()
                    data['totalrechazada'] = totalrechazada = solicitudes.filter(estado=3).count()
                    data['totalpendiente'] = totalneto - (totalrevision + totalaprobada + totalrechazada)
                    data['totalbecaaceptada'] = totalbecaaceptada = solicitudes.filter(estado=2, becaaceptada=2).count()
                    data['totalbecarechazada'] = totalbecarechazada = solicitudes.filter(estado=2,
                                                                                         becaaceptada=3).count()
                    data['totalbecapendiente'] = totalaprobada - (totalbecaaceptada + totalbecarechazada)
                    data['tipobeca'] = BecaTipo.objects.filter(becasolicitud__periodo=periodo, status=True,
                                                               becasolicitud__periodocalifica__isnull=False).order_by(
                        'id').distinct()
                    pc = None
                    periodocalifica_1 = None
                    if BecaSolicitud.objects.values('periodocalifica_id').filter(periodo=periodo, status=True,
                                                                                 periodocalifica__isnull=False).order_by(
                            '-periodocalifica_id').distinct().exists():
                        pc = BecaSolicitud.objects.values('periodocalifica_id').filter(periodo=periodo, status=True,
                                                                                       periodocalifica__isnull=False).order_by(
                            '-periodocalifica_id').distinct()[0]
                        periodocalifica_1 = Periodo.objects.get(pk=pc['periodocalifica_id'])
                    data['periodovalida'] = periodocalifica_1
                    data['btipobeca'] = btipobeca
                    data['bestadoaceptado'] = bestadoaceptado
                    data['bestadoaasignado'] = bestadoaasignado
                    data['bestadobeca'] = bestadobeca
                    return render(request, "adm_becas/listasolicitudes.html", data)
                except Exception as ex:
                    pass

            elif action == 'listaprimernivel':
                try:
                    data['tipobeca'] = tipo = BecaTipo.objects.get(status=True, vigente=True, id=16)
                    configuracion = tipo.becatipoconfiguracion_set.filter(becaperiodo__periodo=periodo).first()
                    totalconf_requisitos = configuracion.requisitosbecas.count() if hasattr(configuracion, 'requisitosbecas') else 0
                    data['title'] = tipo.nombre
                    data['reporte_0'] = obtener_reporte('reporte_estudiantes_primer_nivel')
                    search = None
                    ids = None
                    inscripciones = PreInscripcionBeca.objects.filter(becatipo=tipo).values_list('inscripcion_id', flat=True)
                    matriculas = Matricula.objects.filter(inscripcion__pk__in=inscripciones,
                                                          nivel__periodo=periodo,
                                                          nivelmalla__id=1,
                                                          estado_matricula__in=[2, 3],
                                                          retiradomatricula=False,
                                                          status=True,).exclude(inscripcion__coordinacion__excluir=True)
                    if 's' in request.GET:
                        search = request.GET['s'].strip()
                        matriculas = matriculas.filter(inscripcion__persona__cedula=search)
                    # else:
                    #     matriculas = matriculas.filter(status=True,
                    #                                             nivel__periodo=periodo,
                    #                                             nivelmalla__id=1,
                    #                                             estado_matricula__in=[2, 3],
                    #                                             retiradomatricula=False,
                    #                                             ).exclude(
                    #         inscripcion__coordinacion__excluir=True)
                    matriculas = matriculas.distinct().order_by("inscripcion__persona")

                    # if 's' in request.GET:
                    #     search = request.GET['s'].strip()
                    #     nuevolistado = Matricula.objects.filter(status=True, inscripcion__persona__cedula=search,
                    #                                             nivel__periodo=periodo,
                    #                                             nivelmalla__id=1,
                    #                                             estado_matricula__in=[2, 3],
                    #                                             retiradomatricula=False,
                    #                                             ).exclude(
                    #         inscripcion__coordinacion__excluir=True).distinct().order_by("inscripcion__persona")
                    # else:
                    #     nuevolistado = Matricula.objects.filter(status=True,
                    #                                             nivel__periodo=periodo,
                    #                                             nivelmalla__id=1,
                    #                                             estado_matricula__in=[2, 3],
                    #                                             retiradomatricula=False,
                    #                                             ).exclude(
                    #         inscripcion__coordinacion__excluir=True).distinct().order_by("inscripcion__persona")

                    data['carreras'] = carreras = Carrera.objects.filter(status=True, pk__in=matriculas.values_list(
                        'inscripcion__carrera_id', flat=True).distinct())
                    data['niveles'] = niveles = NivelMalla.objects.filter(status=True, orden=1).order_by('orden')

                    resumen = {}
                    resumen['carreras'] = []
                    idmats = matriculas.values_list('inscripcion_id', flat=True).distinct()
                    estados_matriculas = matriculas.values_list('estado_matricula', flat=True).order_by('estado_matricula').distinct()
                    resumen['estados_matriculas'] = [{'name': dict(ESTADO_MATRICULA)[estadomatr], 'count':matriculas.filter(estado_matricula=estadomatr).count()} for estadomatr in estados_matriculas]
                    resumen['total'] = matriculas.count()
                    resumen['count_renovacion'] = BecaSolicitud.objects.values("id").filter(inscripcion__id__in=idmats, periodo=anterior, becaaceptada=2, becaasignada=2, becatipo=tipo).count()
                    resumen['count_nueva'] = resumen['total'] - resumen['count_renovacion']
                    resumen['totalgeneralrequisitos'] = PreInscripcionBeca.objects.filter(inscripcion_id__in=idmats, periodo=periodo, becatipo=tipo).count()
                    resumen['totalcumplenrequisitos'] = PreInscripcionBeca.objects.filter(inscripcion_id__in=idmats, periodo=periodo, becatipo=tipo, preinscripcionbecarequisito__cumplerequisito=True).annotate(total=Count('preinscripcionbecarequisito__id')).filter(total=totalconf_requisitos).distinct().count()
                    resumen['totalpendientesrequisitos'] = PreInscripcionBeca.objects.filter(inscripcion_id__in=idmats, periodo=periodo, becatipo=tipo, preinscripcionbecarequisito__cumplerequisito__isnull=True).distinct().count()
                    resumen['totalrechazadosrequisitos'] = resumen['totalgeneralrequisitos'] - (resumen['totalcumplenrequisitos'] + resumen['totalpendientesrequisitos'])

                    for carrera in carreras:
                        datacarr = {
                            'name': carrera.nombre,
                            'niveles': []
                        }
                        for nivel in niveles:
                            datacarr['niveles'].append(matriculas.filter(inscripcion__carrera=carrera, nivelmalla=nivel).count())
                        datacarr['total'] = sum(datacarr['niveles'])
                        resumen['carreras'].append(datacarr)
                    data['resumen'] = resumen
                    paging = MiPaginador(matriculas, 25)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['page'] = page
                    data['periodoactual'] = periodo
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['primernivel'] = page.object_list
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['total'] = matriculas.count()
                    return render(request, "adm_becas/listaprimernivel.html", data)
                except Exception as ex:
                    pass

            elif action == 'verarchivos':
                try:
                    data['title'] = u'Verificación de Requisitos y Aprobación de Solicitud'
                    data['id'] = int(request.GET['id'])
                    cabecera = BecaSolicitud.objects.get(pk=int(request.GET['id']))

                    if cabecera.estado == 4 and cabecera.usuario_modificacion_id != persona.usuario_id:
                        return HttpResponseRedirect(
                            "/adm_becas?info=La solicitud de beca ya se encuentra en revisión por otro usuario.")

                    if cabecera.estado != 4:
                        cabecera.estado = 4
                        cabecera.save(request)

                        recorrido = BecaSolicitudRecorrido(solicitud=cabecera,
                                                           observacion="EN REVISION",
                                                           estado=4,
                                                           fecha=datetime.now().date()
                                                           )
                        recorrido.save(request)

                        # Envio de e-mail de notificación al solicitante
                        tituloemail = "Solicitud de Beca en Revisión"
                        tipobeca = cabecera.becatipo.nombre.upper()

                        send_html_mail(tituloemail,
                                       "emails/notificarestadosolicitudbeca.html",
                                       {'sistema': u'SGA - UNEMI',
                                        'fase': 'REV',
                                        'tipobeca': tipobeca,
                                        'fecha': datetime.now().date(),
                                        'hora': datetime.now().time(),
                                        'saludo': 'Estimada' if cabecera.inscripcion.persona.sexo_id == 1 else 'Estimado',
                                        'estudiante': cabecera.inscripcion.persona.nombre_completo_inverso(),
                                        'autoridad2': '',
                                        't': miinstitucion()
                                        },
                                       cabecera.inscripcion.persona.lista_emails_envio(),
                                       [],
                                       cuenta=variable_valor('CUENTAS_CORREOS')[0]
                                       )

                        log(u'Editó estado de solicitud de beca: %s' % persona, request, "edit")

                    data['cabecerasolicitud'] = cabecera
                    data['detallesolicitud'] = detalle = cabecera.becadetallesolicitud_set.filter(status=True).order_by(
                        'requisito')
                    data['totalitems'] = detalle.count()
                    data['totalcumple'] = detalle.filter(cumple=True).count()
                    data['totalarchivocumple'] = detalle.filter(estado=2).count()
                    data['totalarchivonocumple'] = detalle.filter(estado=3).count()
                    # data['documentopersonal'] = cabecera.inscripcion.persona.documentos_personales()
                    return render(request, "adm_becas/verarchivos.html", data)

                except Exception as ex:
                    transaction.set_rollback(True)
                    pass

            elif action == 'detallesolicitud':
                try:
                    data = {}
                    lista = []
                    data['cabecerasolicitud'] = cabecera = BecaSolicitud.objects.get(pk=int(request.GET['id']))
                    if periodo.id < 119:
                        data['detallesolicitud'] = cabecera.becadetallesolicitud_set.filter(status=True).order_by('requisito__id')
                    else:
                        data['detallesolicitud'] = cabecera.becadetallesolicitud_set.filter(status=True, requisito__requisitogeneral__visible=True).order_by('requisito__id')
                    template = get_template("adm_becas/detallesolicitud.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content})
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

            elif action == 'mostrarrecorrido':
                try:
                    data = {}
                    data['solicitud'] = solicitud = BecaSolicitud.objects.get(pk=int(request.GET['id']))
                    data['recorrido'] = solicitud.becasolicitudrecorrido_set.filter(status=True).order_by('id')
                    template = get_template("adm_becas/recorridosolicitud.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content})
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

            elif action == 'subirarchivopendiente':
                try:
                    data = {}
                    data['solicitud'] = solicitud = BecaSolicitud.objects.get(pk=int(request.GET['id']))
                    requisito = BecaRequisitos.objects.get(periodo_id=90, status=True, vigente=True, pk=15)
                    req9 = requisito.nombre
                    requisito = BecaRequisitos.objects.get(periodo_id=90, status=True, vigente=True, pk=16)
                    req10 = requisito.nombre
                    data['req9'] = req9
                    data['req10'] = req10

                    template = get_template("adm_becas/subirarchivopendiente.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content})
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

            elif action == 'probabilidad':
                try:
                    data = {}
                    ins = Inscripcion.objects.get(pk=int(request.GET['id']))
                    data['matricula'] = Matricula.objects.get(inscripcion=ins, nivel__periodo=periodo)
                    data['requisito'] = BecaRequisitos.objects.filter(status=True, vigente='True').order_by('id')
                    data['idper'] = periodo.id
                    data['tipobeca'] = BecaTipo.objects.get(status=True, vigente=True, id=int(request.GET['tipo']))
                    template = get_template("adm_becas/probabilidad.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content})
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

            elif action == 'aprobarrechazar':
                try:
                    solicitud = BecaSolicitud.objects.get(pk=request.GET['id'])
                    return JsonResponse({"result": "ok", "idsolicitud": solicitud.id,
                                         "estado": solicitud.estado,
                                         "observacion": solicitud.observacion})
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": "Error al guardar los datos."})

            elif action == 'aprobarrechazararchivo':
                try:
                    data['title'] = u'Aprobar o Rechazar Evidencia'
                    data['idc'] = int(request.GET['idc'])
                    data['form'] = BecaAprobarArchivoForm
                    data['detalle'] = BecaDetalleSolicitud.objects.get(pk=int(request.GET['id']))
                    template = get_template(
                        "adm_becas/aprobarrechazararchivo.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content, 'title': data['title']})
                except Exception as ex:
                    pass

            elif action == 'resumentotales':
                try:
                    data['title'] = u'Resumen de Total Neto de Solicitudes'

                    modalidades = []
                    datos = []
                    necesidades = []

                    solicitudes = BecaSolicitud.objects.filter(status=True, periodo=periodo).exclude(estado__in=[5, 8])

                    totalsol = solicitudes.count()
                    totalsolapro = solicitudes.filter(estado=2).count()
                    totalbecaacep = solicitudes.filter(estado=2, becaaceptada=2).count()
                    totalbecarecha = solicitudes.filter(estado=2, becaaceptada=3).count()

                    totaldiscapacidad = solicitudes.filter(becatipo_id=19).count()
                    totalcategoriac = solicitudes.filter(becatipo_id=18,
                                                         inscripcion__persona__fichasocioeconomicainec__grupoeconomico__codigo='C-').count()
                    totalcategoriad = solicitudes.filter(becatipo_id=18,
                                                         inscripcion__persona__fichasocioeconomicainec__grupoeconomico__codigo='D').count()

                    totaldiscapacidadapro = solicitudes.filter(becatipo_id=19, estado=2).count()
                    totalcategoriacapro = solicitudes.filter(becatipo_id=18, estado=2,
                                                             inscripcion__persona__fichasocioeconomicainec__grupoeconomico__codigo='C-').count()
                    totalcategoriadapro = solicitudes.filter(becatipo_id=18, estado=2,
                                                             inscripcion__persona__fichasocioeconomicainec__grupoeconomico__codigo='D').count()

                    totaltabletyplandatos = solicitudes.filter(becasolicitudnecesidad__necesidad=1).count()
                    totalplandatos = solicitudes.filter(becasolicitudnecesidad__necesidad=2).count()
                    totaltabletyplanapro = solicitudes.filter(estado=2, becasolicitudnecesidad__necesidad=1).count()
                    totaltabletyplanbecaacep = solicitudes.filter(estado=2, becaaceptada=2,
                                                                  becasolicitudnecesidad__necesidad=1).count()
                    totaltabletyplanbecarecha = solicitudes.filter(estado=2, becaaceptada=3,
                                                                   becasolicitudnecesidad__necesidad=1).count()
                    totalplandatosapro = solicitudes.filter(estado=2, becasolicitudnecesidad__necesidad=2).count()
                    totalplandatosbecaacep = solicitudes.filter(estado=2, becaaceptada=2,
                                                                becasolicitudnecesidad__necesidad=2).count()
                    totalplandatosbecarecha = solicitudes.filter(estado=2, becaaceptada=3,
                                                                 becasolicitudnecesidad__necesidad=2).count()

                    totalpresencial = solicitudes.filter(inscripcion__modalidad_id=1).count()
                    totalsemipresencial = solicitudes.filter(inscripcion__modalidad_id=2).count()
                    totalenlinea = solicitudes.filter(inscripcion__modalidad_id=3).count()

                    totalpresencialapro = solicitudes.filter(estado=2, inscripcion__modalidad_id=1).count()
                    totalpresencialbecaacep = solicitudes.filter(estado=2, becaaceptada=2,
                                                                 inscripcion__modalidad_id=1).count()
                    totalpresencialbecarecha = solicitudes.filter(estado=2, becaaceptada=3,
                                                                  inscripcion__modalidad_id=1).count()
                    totalsemipresencialapro = solicitudes.filter(estado=2, inscripcion__modalidad_id=2).count()
                    totalsemipresencialbecaacep = solicitudes.filter(estado=2, becaaceptada=2,
                                                                     inscripcion__modalidad_id=2).count()
                    totalsemipresencialbecarecha = solicitudes.filter(estado=2, becaaceptada=3,
                                                                      inscripcion__modalidad_id=2).count()
                    totalenlineaapro = solicitudes.filter(estado=2, inscripcion__modalidad_id=3).count()
                    totalenlineabecaacep = solicitudes.filter(estado=2, becaaceptada=2,
                                                              inscripcion__modalidad_id=3).count()
                    totalenlineabecarecha = solicitudes.filter(estado=2, becaaceptada=3,
                                                               inscripcion__modalidad_id=3).count()

                    modalidades.append(
                        ['Total Modalidad Presencial: ', totalpresencial, totalpresencialapro, totalpresencialbecaacep,
                         totalpresencialbecarecha])
                    modalidades.append(
                        ['Total Modalidad Semipresencial: ', totalsemipresencial, totalsemipresencialapro,
                         totalsemipresencialbecaacep, totalsemipresencialbecarecha])
                    modalidades.append(
                        ['Total Modalidad En Línea: ', totalenlinea, totalenlineaapro, totalenlineabecaacep,
                         totalenlineabecarecha])

                    datos.append(
                        ['Total Grupo Socioeconómico <strong>C-</strong> :', totalcategoriac, totalcategoriacapro])
                    datos.append(
                        ['Total Grupo Socioeconómico <strong>D</strong> :', totalcategoriad, totalcategoriadapro])
                    datos.append(['Total por Discapacidad :', totaldiscapacidad, totaldiscapacidadapro])

                    necesidades.append(['TABLET+PLAN DE DATOS :', totaltabletyplandatos, totaltabletyplanapro,
                                        totaltabletyplanbecaacep, totaltabletyplanbecarecha])
                    necesidades.append(['PLAN DE DATOS :', totalplandatos, totalplandatosapro, totalplandatosbecaacep,
                                        totalplandatosbecarecha])

                    data['modalidades'] = modalidades
                    data['datos'] = datos
                    data['necesidades'] = necesidades
                    data['totalsol'] = totalsol
                    data['totalsolapro'] = totalsolapro
                    data['totalbecaacep'] = totalbecaacep
                    data['totalbecarecha'] = totalbecarecha
                    template = get_template("adm_becas/resumentotalsolicitud.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content, 'title': data['title']})
                except Exception as ex:
                    pass

            elif action == 'resumencupos':
                try:
                    data['title'] = u'Resumen Cupos Asignados y Disponibles'

                    cupopresencial = 1177
                    cupoenlinea = 450

                    becas = BecaAsignacion.objects.filter(status=True, solicitud__periodo=periodo).exclude(estadobeca=3)
                    becaspresencial = becas.filter(solicitud__inscripcion__modalidad_id__in=[1, 2])
                    becasenlinea = becas.filter(solicitud__inscripcion__modalidad_id=3)

                    total = becas.count()
                    totalpresencial = becaspresencial.count()
                    totalenlinea = becasenlinea.count()

                    totalpresencialval = becaspresencial.filter(estadorevisioncontrato=2).count()
                    totalenlineaval = becasenlinea.filter(estadorevisioncontrato=2).count()

                    datos = []

                    datos.append(['Presencial/Semipresencial:', cupopresencial, totalpresencial,
                                  cupopresencial - totalpresencial, totalpresencialval,
                                  cupopresencial - totalpresencialval])
                    datos.append(['En Línea:', cupoenlinea, totalenlinea, cupoenlinea - totalenlinea, totalenlineaval,
                                  cupoenlinea - totalenlineaval])

                    data['datos'] = datos
                    data['totalcupos'] = totalcupos = cupopresencial + cupoenlinea
                    data['totalbecas'] = total
                    data['totaldisponibles'] = totalcupos - total
                    data['totalvalidado'] = totalval = totalpresencialval + totalenlineaval
                    data['totaldisponiblesvalidado'] = totalcupos - totalval
                    template = get_template("adm_becas/resumencupos.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content, 'title': data['title']})
                except Exception as ex:
                    pass

            elif action == 'resumenbecasasignadas':
                try:
                    data['title'] = u'Resumen Becas Asignadas'

                    becas1 = BecaAsignacion.objects.filter(status=True, solicitud__periodo=periodo, tipo=1,
                                                           estadorevisioncontrato=2).exclude(estadobeca=3)
                    becas2 = BecaAsignacion.objects.filter(status=True, solicitud__periodo=periodo, tipo=2).exclude(
                        estadobeca=3)

                    becas = becas1 | becas2
                    tiposbecas = becas.values("solicitud__becatipo", "solicitud__becatipo__nombre").distinct().order_by(
                        'solicitud__becatipo__nombre')
                    datos = []
                    total1vez = totalrenova = 0

                    for tipo in tiposbecas:
                        idtipo = tipo['solicitud__becatipo']
                        ntipo = tipo['solicitud__becatipo__nombre'].upper()
                        total = becas.filter(solicitud__becatipo_id=idtipo).aggregate(total=Count('id'))['total']
                        totalnuevas = becas.filter(solicitud__becatipo_id=idtipo, tipo=1).aggregate(total=Count('id'))[
                            'total']
                        totalrenovacion = \
                        becas.filter(solicitud__becatipo_id=idtipo, tipo=2).aggregate(total=Count('id'))['total']

                        datos.append([ntipo, totalnuevas, totalrenovacion, total])
                        total1vez += totalnuevas
                        totalrenova += totalrenovacion

                    data['datos'] = datos
                    data['total1vez'] = total1vez
                    data['totalrenovacion'] = totalrenova
                    data['totalbeca'] = total1vez + totalrenova
                    template = get_template("adm_becas/resumenbecassaignadas.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content, 'title': data['title']})
                except Exception as ex:
                    pass

            elif action == 'anularbeca':
                try:
                    data['title'] = u'Anular Beca'
                    data['form'] = BecaSolicitudAnulacionForm()
                    data['id'] = id = request.GET.get('id','')
                    if id:
                        data['beca'] = BecaAsignacion.objects.get(pk=int(request.GET['id']))
                    template = get_template("adm_becas/anularbeca.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content, 'title': data['title']})
                except Exception as ex:
                    pass

            elif action == 'actualizarpagado':
                try:
                    if not SolicitudPagoBecaDetalle.objects.values("id").filter(asignacion__solicitud__periodo=periodo,
                                                                                pagado=False).exists():
                        return JsonResponse(
                            {"result": "bad",
                             "mensaje": u"No existen registros de becas pendientes de actualizar a PAGADO para el periodo seleccionado"})

                    data['title'] = u'Actualizar Becas a Estado Pagado'
                    data['form'] = ImportarArchivoPagoBecaXLSForm()
                    template = get_template("adm_becas/pagobeca.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content, 'title': data['title']})
                except Exception as ex:
                    pass

            elif action == 'actualizaracreditado':
                try:
                    if not SolicitudPagoBecaDetalle.objects.values("id").filter(asignacion__solicitud__periodo=periodo,
                                                                                pagado=True,
                                                                                acreditado=False).exists():
                        return JsonResponse(
                            {"result": "bad",
                             "mensaje": u"No existen registros de becas pendientes de actualizar a ACREDITADO para el periodo seleccionado"})

                    data['title'] = u'Actualizar Becas a Estado Acreditado'
                    data['form'] = ImportarArchivoPagoBecaXLSForm()
                    template = get_template("adm_becas/acreditabeca.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content, 'title': data['title']})
                except Exception as ex:
                    pass

            elif action == 'rechazarsolicitudonline':
                try:
                    data['title'] = u'Rechazar Solicitud Estudiantes Modalidad En Línea'
                    data['id'] = int(request.GET['id'])
                    data['form'] = BecaSolicitudRechazaOnLineForm
                    data['solicitud'] = BecaSolicitud.objects.get(pk=int(request.GET['id']))
                    data[
                        'observacion'] = 'LAS SOLICITUDES DE BECAS PARA EL PERIODO MAYO - SEPTIEMBRE 2020 V2 SÓLO APLICAN PARA ESTUDIANTES DE LAS MODALIDADES PRESENCIAL Y SEMIPRESENCIAL'
                    template = get_template("adm_becas/rechazarsolicitudonline.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content, 'title': data['title']})
                except Exception as ex:
                    pass

            elif action == 'validarcedula':
                try:
                    data['title'] = u'Revisar/Validar Documentos'
                    data['idb'] = int(request.GET['idb'])
                    data['beca'] = beca = BecaAsignacion.objects.get(pk=int(request.GET['idb']))
                    estudiante = beca.solicitud.inscripcion.persona
                    representante = estudiante.personaextension_set.all()[0]
                    data['estadodocumento'] = request.GET['estadodocumento']
                    #data['estadocontrato'] = request.GET['estadocontrato']
                    data['representante'] = representante

                    if beca.personarevisadocumento:
                        if beca.personarevisadocumento != persona:
                            return JsonResponse({"result": "bad",
                                                 "mensaje": "Los documentos de la beca están siendo revisados por otro usuario."})
                    else:
                        beca.personarevisadocumento = persona
                        beca.save(request)

                    solicitante = beca.solicitud.inscripcion.persona
                    data['permite_modificar'] = permite_modificar = False
                    documentos = solicitante.documentos_personales()
                    cuentabancaria = solicitante.cuentabancaria()
                    perilinscripcion = beca.solicitud.inscripcion.persona.mi_perfil()

                    if documentos:
                        if beca.solicitud.becatipo.id == 16:
                            if documentos.estadocedula == 1 or documentos.estadocedularepresentantesol == 1 or documentos.estadopapeleta == 1 or documentos.estadopapeletarepresentantesol == 1 or documentos.estadoactagrado == 1:
                                data['permite_modificar'] = permite_modificar = True
                                data['estadocedula1'] = documentos.estadocedula
                                data['estadocert1'] = documentos.estadopapeleta
                                data['estadocedula2'] = documentos.estadocedularepresentantesol
                                data['estadocert2'] = documentos.estadopapeletarepresentantesol
                                data['estadoacta'] = documentos.estadoactagrado
                        elif beca.solicitud.becatipo.id == 21:
                            if documentos.estadocedula == 1 or documentos.estadocedularepresentantesol == 1 or documentos.estadopapeleta == 1 or documentos.estadopapeletarepresentantesol == 1 or perilinscripcion.estadoarchivoraza == 1:
                                data['permite_modificar'] = permite_modificar = True
                                data['estadocedula1'] = documentos.estadocedula
                                data['estadocert1'] = documentos.estadopapeleta
                                data['estadocedula2'] = documentos.estadocedularepresentantesol
                                data['estadocert2'] = documentos.estadopapeletarepresentantesol

                        elif beca.solicitud.becatipo.id == 22 and solicitante.ecuatoriano_vive_exterior():
                            if documentos.estadocedula == 1 or documentos.estadocedularepresentantesol == 1 or documentos.estadopapeleta == 1 or documentos.estadopapeletarepresentantesol == 1 or documentos.estadoserviciosbasico == 1:
                                data['permite_modificar'] = permite_modificar = True
                                data['estadocedula1'] = documentos.estadocedula
                                data['estadocert1'] = documentos.estadopapeleta
                                data['estadocedula2'] = documentos.estadocedularepresentantesol
                                data['estadocert2'] = documentos.estadopapeletarepresentantesol
                        else:
                            if documentos.estadocedula == 1 or documentos.estadocedularepresentantesol == 1 or documentos.estadopapeleta == 1 or documentos.estadopapeletarepresentantesol == 1:
                                data['permite_modificar'] = permite_modificar = True
                                data['estadocedula1'] = documentos.estadocedula
                                data['estadocert1'] = documentos.estadopapeleta
                                data['estadocedula2'] = documentos.estadocedularepresentantesol
                                data['estadocert2'] = documentos.estadopapeletarepresentantesol

                    data['documentos'] = documentos
                    data['cuentabancaria'] = cuentabancaria
                    data['perilinscripcion'] = perilinscripcion
                    if beca.solicitud.periodo_id >= 119:
                        if beca.solicitud.becatipo.id != 16:
                            template = get_template("adm_becas/validarcedula_new.html")
                        else:
                            template = get_template("adm_becas/validarcedulabpn.html")
                    else:
                        if beca.solicitud.becatipo.id != 16:
                            template = get_template("adm_becas/validarcedula.html")
                        else:
                            template = get_template("adm_becas/validarcedulabpn.html")

                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content, 'title': data['title'],
                                         'permite_modificar': permite_modificar})
                except Exception as ex:
                    pass

            elif action == 'validarcontrato':
                try:
                    data['title'] = u'Revisar/Validar Contrato de Beca'
                    data['idb'] = int(request.GET['idb'])
                    data['beca'] = beca = BecaAsignacion.objects.get(pk=int(request.GET['idb']))
                    data['estadodocumento'] = request.GET['estadodocumento']
                    data['estadocontrato'] = request.GET['estadocontrato']
                    data['documentos'] = beca.solicitud.inscripcion.persona.documentos_personales()

                    if beca.personarevisacontrato:
                        if beca.personarevisacontrato != persona:
                            return JsonResponse({"result": "bad",
                                                 "mensaje": "El contrato de beca está siendo revisado por otro usuario."})
                    else:
                        beca.personarevisacontrato = persona
                        beca.save(request)

                    if beca.estadorevisioncontrato == 2 or beca.estadorevisioncontrato == 3:
                        data['permite_modificar'] = False
                    else:
                        data['permite_modificar'] = True

                    template = get_template("adm_becas/validarcontrato.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content, 'title': data['title']})
                except Exception as ex:
                    pass

            elif action == 'validaradquisicion':
                try:
                    data['title'] = u'Revisar/Validar Cumplimiento Aquisición'
                    data['tipovalida'] = 'ADQ'
                    data['idb'] = int(request.GET['idb'])
                    data['beca'] = beca = BecaAsignacion.objects.get(pk=int(request.GET['idb']))
                    data['estadodocumento'] = request.GET['estadodocumento']
                    data['estadocontrato'] = request.GET['estadocontrato']
                    data['estadocumplimiento'] = request.GET['estadocumplimiento']
                    data['estadocomprobante'] = request.GET['estadocomprobante']

                    # data['comprobante'] = comprobante = beca.becacomprobanteventa_set.filter(status=True)[0]
                    revisioncomprobante = beca.becacomprobanterevision_set.filter(status=True)[0]
                    data['valorminimo'] = null_to_decimal(beca.montobeneficio * 0.70, 2)
                    data['comprobantes'] = comprobantes = revisioncomprobante.becacomprobanteventa_set.filter(
                        status=True).order_by('id')
                    data['totalcomprobantes'] = comprobantes.aggregate(total=Sum('total'))['total']
                    data['revisionid'] = revisioncomprobante.id
                    if revisioncomprobante.personarevisadbu:
                        if revisioncomprobante.personarevisadbu != persona:
                            return JsonResponse({"result": "bad",
                                                 "mensaje": "El comprobante de venta está siendo revisado por otro usuario."})
                    else:
                        revisioncomprobante.personarevisadbu = persona
                        revisioncomprobante.estadorevisiondbu = 5
                        revisioncomprobante.save(request)

                    if revisioncomprobante.estadorevisiondbu == 2 or revisioncomprobante.estadorevisiondbu == 3 or revisioncomprobante.estadorevisiondbu == 6:
                        data['permite_modificar'] = False
                        rp = "N"
                    else:
                        data['permite_modificar'] = True
                        rp = "S"

                    data['puede_eliminar'] = True if comprobantes.count() > 1 else False

                    template = get_template("adm_becas/validaradquisicion.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content, 'rp': rp, 'title': data['title']})
                except Exception as ex:
                    pass

            elif action == 'detallecomprobantes':
                try:
                    revisioncomprobante = BecaComprobanteRevision.objects.get(pk=int(request.GET['idrc']))
                    comprobantes = revisioncomprobante.becacomprobanteventa_set.filter(status=True).order_by('id')
                    data['comprobantes'] = comprobantes
                    template = get_template("adm_becas/detallecomprobantes.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content})
                except Exception as ex:
                    pass

            elif action == 'editarcomprobanteventa':
                try:
                    data['title'] = u'Revisar/Validar Comprobante de Venta (Bienestar)' if request.GET[
                                                                                               'tva'] == 'ADQ' else u'Revisar/Validar Comprobante de Venta (Financiero)'
                    data['idc'] = int(request.GET['idc'])

                    comprobante = BecaComprobanteVenta.objects.get(pk=request.GET['idc'])

                    form = BecaComprobanteVentaValidaForm(initial={'rucproveedor': comprobante.rucproveedor,
                                                                   'total': comprobante.total})

                    form.estados_validar()
                    data['form'] = form

                    template = get_template("adm_becas/editcomprobante.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content, 'title': data['title']})
                except Exception as ex:
                    pass

            elif action == 'eliminarcomprobanteventa':
                try:
                    data['title'] = u'Eliminar Comprobante de Venta'
                    data['idc'] = int(request.GET['idc'])

                    comprobante = BecaComprobanteVenta.objects.get(pk=request.GET['idc'])

                    form = BecaComprobanteEliminarForm()

                    data['form'] = form

                    template = get_template("adm_becas/delcomprobante.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content, 'title': data['title']})
                except Exception as ex:
                    pass

            # elif action == 'validarcomprobante':
            #     try:
            #         data['title'] = u'Revisar/Validar Comprobante de Venta'
            #         data['idb'] = int(request.GET['idb'])
            #         data['beca'] = beca = BecaAsignacion.objects.get(pk=int(request.GET['idb']))
            #         data['estadodocumento'] = request.GET['estadodocumento']
            #         data['estadocontrato'] = request.GET['estadocontrato']
            #         data['estadocumplimiento'] = request.GET['estadocumplimiento']
            #         data['estadocomprobante'] = request.GET['estadocomprobante']
            #         data['comprobante'] = comprobante = beca.becacomprobanteventa_set.filter(status=True)[0]
            #         data['valorminimo'] = null_to_decimal(beca.montobeneficio * 0.70, 2)
            #
            #         if comprobante.personarevisafin:
            #             if comprobante.personarevisafin != persona:
            #                 return JsonResponse({"result": "bad",
            #                                      "mensaje": "El comprobante de venta está siendo revisado por otro usuario."})
            #         else:
            #             comprobante.personarevisafin = persona
            #             comprobante.estadorevisionfin = 5
            #             comprobante.save(request)
            #
            #         if comprobante.estadorevisionfin == 2 or comprobante.estadorevisionfin == 3 or comprobante.estadorevisionfin == 6:
            #             data['permite_modificar'] = False
            #         else:
            #             data['permite_modificar'] = True
            #
            #         template = get_template("adm_becas/validarcomprobante.html")
            #         json_content = template.render(data)
            #         return JsonResponse({"result": "ok", 'data': json_content, 'title': data['title']})
            #     except Exception as ex:
            #         pass

            elif action == 'validarcomprobante':
                try:
                    data['title'] = u'Revisar/Validar Comprobente'
                    data['tipovalida'] = 'CVTA'
                    data['idb'] = int(request.GET['idb'])
                    data['beca'] = beca = BecaAsignacion.objects.get(pk=int(request.GET['idb']))
                    data['estadodocumento'] = request.GET['estadodocumento']
                    data['estadocontrato'] = request.GET['estadocontrato']
                    data['estadocumplimiento'] = request.GET['estadocumplimiento']
                    data['estadocomprobante'] = request.GET['estadocomprobante']

                    # data['comprobante'] = comprobante = beca.becacomprobanteventa_set.filter(status=True)[0]
                    revisioncomprobante = beca.becacomprobanterevision_set.filter(status=True)[0]
                    data['valorminimo'] = null_to_decimal(beca.montobeneficio * 0.70, 2)
                    data['comprobantes'] = comprobantes = revisioncomprobante.becacomprobanteventa_set.filter(
                        status=True).order_by('id')
                    data['totalcomprobantes'] = comprobantes.aggregate(total=Sum('total'))['total']
                    data['revisionid'] = revisioncomprobante.id
                    if revisioncomprobante.personarevisafin:
                        if revisioncomprobante.personarevisafin != persona:
                            return JsonResponse({"result": "bad",
                                                 "mensaje": "El comprobante de venta está siendo revisado por otro usuario."})
                    else:
                        revisioncomprobante.personarevisafin = persona
                        revisioncomprobante.estadorevisionfin = 5
                        revisioncomprobante.save(request)

                    if revisioncomprobante.estadorevisionfin == 2 or revisioncomprobante.estadorevisionfin == 3 or revisioncomprobante.estadorevisionfin == 6:
                        data['permite_modificar'] = False
                        rp = "N"
                    else:
                        data['permite_modificar'] = True
                        rp = "S"

                    template = get_template("adm_becas/validaradquisicion.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content, 'rp': rp, 'title': data['title']})
                except Exception as ex:
                    pass

            elif action == 'representantesolidario':
                try:
                    data['title'] = u'Actualizar Representante Solidario'
                    data['id'] = int(request.GET['id'])
                    data['estadodocumento'] = request.GET['estadodocumento']
                    beca = BecaAsignacion.objects.get(pk=int(request.GET['id']))
                    estudiante = beca.solicitud.inscripcion.persona
                    representante = estudiante.personaextension_set.all()[0]

                    data['form'] = RepresentanteSolidarioForm(initial={'cedula': representante.cedularepsolidario,
                                                                       'nombre': representante.nombresrepsolidario,
                                                                       'apellido1': representante.apellido1repsolidario,
                                                                       'apellido2': representante.apellido2repsolidario
                                                                       })

                    template = get_template("adm_becas/representantesolidario.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content, 'title': data['title']})
                except Exception as ex:
                    pass

            elif action == 'asignarbeca2':
                try:
                    data['title'] = u'Asignar Beca'
                    data['id'] = int(request.GET['id'])
                    data['form'] = BecaAsignacion2Form
                    data['solicitud'] = solicitud = BecaSolicitud.objects.get(pk=int(request.GET['id']))

                    beneficios = solicitud.becasolicitudnecesidad_set.all()[0]
                    beneficio = beneficios.get_necesidad_display()
                    beneficioid = beneficios.necesidad
                    data['beneficio'] = beneficio
                    data['beneficioid'] = beneficioid
                    template = get_template("adm_becas/asignarbeca2.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content, 'title': data['title']})
                except Exception as ex:
                    pass

            elif action == 'becautilizacion':
                try:
                    data['title'] = u'JUSTIFICATIVOS PARA LOS DESEMBOLSOS DE LA BECA'
                    data['id'] = id = int(request.GET['id'])
                    data['asignacion'] = asignacion = \
                    BecaAsignacion.objects.filter(status=True, solicitud__inscripcion_id=id,
                                                  solicitud__periodo=periodo)[0]
                    utilizacion = BecaUtilizacion.objects.filter(status=True, vigente=True)
                    data['utilizacion'] = utilizacion
                    if BecaDetalleUtilizacion.objects.values('id').filter(status=True, asignacion=asignacion).exists():
                        data['detalle'] = BecaDetalleUtilizacion.objects.filter(status=True, asignacion=asignacion)
                    return render(request, "adm_becas/becautilizacionalumno.html", data)
                except Exception as ex:
                    pass

            elif action == 'aprobarrechazarutilizacion':
                try:
                    data['title'] = u'Aprobar o Rechazar Evidencia Utilización'
                    data['idutilizacion'] = idutilizacion = int(request.GET['idutilizacion'])
                    data['idinscripcion'] = int(request.GET['idinscripcion'])
                    data['form'] = BecaAprobarArchivoUtilizacionForm
                    data['detalle'] = BecaDetalleUtilizacion.objects.get(pk=int(idutilizacion))
                    template = get_template("adm_becas/aprobarrechazararchivoutilizacion.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content, 'title': data['title']})
                except Exception as ex:
                    pass

            elif action == 'asignarbeca':
                try:
                    data['title'] = u'ASIGNAR BECA'
                    data['idsolicitud'] = idsolicitud = int(request.GET['idsolicitud'])
                    cabecera = BecaSolicitud.objects.get(pk=idsolicitud)
                    form = BecaAsignacionForm(
                        initial={'solicitud': cabecera.inscripcion.persona.nombre_completo_inverso()})
                    form.bloquear()
                    data['form'] = form
                    template = get_template("alu_becas/asignarbeca.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content, 'title': data['title']})
                except Exception as ex:
                    pass

            elif action == 'detallerquisito':
                try:
                    data = {}
                    idrequisito = int(request.GET['idr'])
                    data['requisito'] = BecaRequisitos.objects.get(id=idrequisito)
                    data['idsolicitud'] = idsolicitud = int(request.GET['ids'])
                    data['detalle'] = BecaDetalleSolicitud.objects.filter(status=True, solicitud_id=idsolicitud,
                                                                          requisito_id=idrequisito)
                    template = get_template("adm_becas/detallerequisito.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

            # elif action == 'subirarchivoutilizacion':
            #     try:
            #         data['title'] = u'Subir Evidencia Utilización'
            #         data['idasignacion'] = idasignacion = request.GET['idasignacion']
            #         data['idutilizacion'] = idutilizacion = request.GET['idutilizacion']
            #         data['idinscripcion'] = request.GET['idinscripcion']
            #         # utilizacion=BecaUtilizacion.objects.get(id=idutilizacion)
            #         # asignacion=BecaAsignacion.objects.get(id=idasignacion)
            #         form = BecaDetalleUtilizacionForm(initial={'asignacion': idasignacion,
            #                                                    'utilizacion': idutilizacion})
            #         form.bloquear()
            #         data['form'] = form
            #         template = get_template("alu_becas/add_evidencia_utilizacion.html")
            #         json_content = template.render(data)
            #         return JsonResponse({"result": "ok", 'data': json_content, 'title': data['title']})
            #     except Exception as ex:
            #         pass

            elif action == 'deletesolicitud':
                try:
                    data['title'] = u'Eliminar Solicitud'
                    data['solicitud'] = BecaSolicitud.objects.get(pk=int(request.GET['id']), status=True)
                    return render(request, "adm_becas/deletesolicitud.html", data)
                except Exception as ex:
                    pass

            elif action == 'editbeca':
                try:
                    data['title'] = u'Editar Beca'
                    data['id'] = int(request.GET['id'])
                    data['becaasignacion'] = becaasignacion = BecaAsignacion.objects.get(pk=int(request.GET['id']),
                                                                                         status=True)
                    form = BecaAsignacionForm(initial={
                        'solicitud': becaasignacion.solicitud.inscripcion.persona.nombre_completo_inverso(),
                        'montobeneficio': becaasignacion.montobeneficio,
                        'montomensual': becaasignacion.montomensual,
                        'cantidadmeses': becaasignacion.cantidadmeses,
                        'tipo': becaasignacion.tipo,
                        'grupopago': becaasignacion.grupopago,
                    })
                    form.bloquear()
                    data['form'] = form
                    template = get_template("adm_becas/editbeca.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content, 'title': data['title']})
                except Exception as ex:
                    pass

            elif action == 'deletebecaasignacion':
                try:
                    data['title'] = u'Eliminar Beca'
                    data['becaasignacion'] = BecaAsignacion.objects.get(pk=int(request.GET['id']), status=True)
                    return render(request, "adm_becas/deletebecaasignacion.html", data)
                except Exception as ex:
                    pass

            elif action == 'importarbecados':
                try:
                    data['title'] = u'Importar Becados'
                    form = ImportarBecaForm()
                    data['form'] = form
                    return render(request, "adm_becas/importarbecados.html", data)
                except Exception as ex:
                    pass

            elif action == 'addbecaperiodo':
                try:
                    data['title'] = u'Nuevo Periodo Beca'
                    data['periodoselec'] = periodoselec = periodosesion
                    form = BecaPeriodoForm(initial={'periodo': periodoselec, 'vigente': True})
                    form.deshabilitar()
                    data['form'] = form
                    return render(request, "adm_becas/addbecaperiodo.html", data)
                except Exception as ex:
                    pass

            elif action == 'editbecaperiodo':
                try:
                    data['title'] = u'Editar Periodo Beca'
                    id = int(request.GET['id'])
                    becaperiodo = BecaPeriodo.objects.get(pk=id)
                    data['periodoselec'] = periodoselec = periodosesion
                    form = BecaPeriodoForm(initial={'periodo': becaperiodo.periodo,
                                                    'fechainiciosolicitud': becaperiodo.fechainiciosolicitud,
                                                    'fechafinsolicitud': becaperiodo.fechafinsolicitud,
                                                    'fechainicioimprimircontrato': becaperiodo.fechainicioimprimircontrato.date() if becaperiodo.fechainicioimprimircontrato else None,
                                                    'horainicioimprimircontrato': becaperiodo.fechainicioimprimircontrato.time() if becaperiodo.fechainicioimprimircontrato else None,
                                                    'fechafinimprimircontrato': becaperiodo.fechafinimprimircontrato.date() if becaperiodo.fechainicioimprimircontrato else None,
                                                    'horafinimprimircontrato': becaperiodo.fechafinimprimircontrato.time() if becaperiodo.fechainicioimprimircontrato else None,
                                                    'fechainiciovalidaciondocumento': becaperiodo.fechainiciovalidaciondocumento,
                                                    'fechafinvalidaciondocumento': becaperiodo.fechafinvalidaciondocumento,
                                                    'fechainiciovalidacioncontrato': becaperiodo.fechainiciovalidacioncontrato,
                                                    'fechafinvalidacioncontrato': becaperiodo.fechafinvalidacioncontrato,
                                                    'fechainicioactualizarcertificariobancario': becaperiodo.fechainicioactualizarcertificariobancario,
                                                    'fechafinactualizarcertificariobancario': becaperiodo.fechafinactualizarcertificariobancario,
                                                    'vigente': becaperiodo.vigente,
                                                    'obligadosubircomprobante': becaperiodo.obligadosubircomprobante,
                                                    'limitebecados': becaperiodo.limitebecados,
                                                    'nivelesmalla': becaperiodo.nivelesmalla.filter(status=True),
                                                    })
                    form.deshabilitar()
                    data['form'] = form
                    data['becaperiodo'] = becaperiodo
                    return render(request, "adm_becas/editbecaperiodo.html", data)
                except Exception as ex:
                    pass

            elif action == 'listaperiodos':
                try:
                    data['title'] = u'Gestión de Periodos'
                    search = None
                    ids = None
                    # periodosbeca = BecaSolicitud.objects.filter(status=True).values_list('periodo', flat=True).distinct()
                    # periodos = Periodo.objects.filter(status=True, id__in=periodosbeca)
                    periodos = BecaPeriodo.objects.filter(status=True).order_by('id')
                    if 's' in request.GET:
                        search = request.GET['s']
                        if search.isdigit():
                            periodos = periodos.filter(pk=search)
                        else:
                            periodos = periodos.filter(Q(periodo__nombre__icontains=search))
                    if 'id' in request.GET:
                        ids = request.GET['id']
                        periodos = periodos.filter(id=ids)
                    vigente = 0
                    if 'vigente' in request.GET and int(request.GET['vigente']) > 0:
                        vigente = int(request.GET['vigente'])
                        periodos = periodos.filter(vigente=int(request.GET['vigente']) == 1)
                    paging = MiPaginador(periodos, 25)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['page'] = page
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['periodosbecas'] = page.object_list
                    data['vigente'] = vigente
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    return render(request, "adm_becas/listaperiodos.html", data)
                except Exception as ex:
                    pass

            elif action == 'listabecatipoconfiguracion':
                try:
                    idpb = None
                    if 'idpb' in request.GET and request.GET['idpb']:
                        idpb = request.GET['idpb']
                        if not BecaPeriodo.objects.filter(pk=idpb, status=True).exists():
                            raise NameError(u"No se encontro el periodo de beca")
                    else:
                        raise NameError(u"No se encontro el periodo de beca")
                    becaperiodo = BecaPeriodo.objects.get(pk=idpb)
                    data['title'] = u'Detalle de tipos de becas del periodo %s' % becaperiodo.periodo.nombre
                    search = None
                    ids = None
                    listados = BecaTipoConfiguracion.objects.filter(becaperiodo=becaperiodo, status=True).order_by(
                        'becatipo__nombre')
                    if 's' in request.GET:
                        search = request.GET['s']
                        if search.isdigit():
                            listados = listados.filter(pk=search, status=True).order_by('id')
                        else:
                            listados = listados.filter(Q(becaperiodo__periodo__nombre__icontains=search) |
                                                       Q(becatipo__nombre__icontains=search) |
                                                       Q(becatipo__nombrecorto__icontains=search),
                                                       status=True).order_by('id')
                    if 'id' in request.GET:
                        ids = request.GET['id']
                        listados = listados.filter(id=ids, status=True).order_by('id')

                    paging = MiPaginador(listados, 10)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['page'] = page
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['listados'] = page.object_list
                    data['becaperiodo'] = becaperiodo
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    return render(request, "adm_becas/configuracionbecatipoperiodo.html", data)
                except Exception as ex:
                    pass

            elif action == 'addbecado':
                try:
                    data['title'] = u'Agregar Becado'
                    form = BecaAsignacionManualForm()
                    data['form'] = form
                    return render(request, "adm_becas/addbecado.html", data)
                except Exception as ex:
                    pass

            elif action == 'addBecaTipoConfiguracion':
                try:
                    idpb = None
                    if 'idpb' in request.GET and request.GET['idpb']:
                        idpb = request.GET['idpb']
                        if not BecaPeriodo.objects.filter(pk=idpb, status=True).exists():
                            raise NameError(u"No se encontro el periodo de beca")
                    else:
                        raise NameError(u"No se encontro el periodo de beca")
                    becaperiodo = BecaPeriodo.objects.get(pk=idpb)
                    data['title'] = u'Agregar configuración del periodo %s' % becaperiodo.periodo.nombre
                    data['becaperiodo'] = becaperiodo
                    form = BecaTipoConfiguracionForm()
                    form.adicionar(becaperiodo)
                    data['form'] = form
                    return render(request, "adm_becas/addbecatipoconfiguracion.html", data)
                except Exception as ex:
                    pass

            elif action == 'editBecaTipoConfiguracion':
                try:
                    idpb = None
                    if 'idpb' in request.GET and request.GET['idpb']:
                        idpb = request.GET['idpb']
                        if not BecaPeriodo.objects.filter(pk=idpb, status=True).exists():
                            raise NameError(u"No se encontro el periodo de beca")
                    else:
                        raise NameError(u"No se encontro el periodo de beca")
                    becaperiodo = BecaPeriodo.objects.get(pk=idpb)
                    x = BecaTipoConfiguracion.objects.get(pk=int(request.GET['id']))
                    data['title'] = u'Editar configuración del periodo %s' % becaperiodo.periodo.nombre
                    data['becaperiodo'] = becaperiodo
                    data['becatipoconfiguracion'] = x
                    # form = BecaTipoConfiguracionForm(initial={'becatipo': x.becatipo,
                    #                                           'becamonto': x.becamonto,
                    #                                           'becameses': x.becameses,
                    #                                           'becamensual': x.monto_x_mes()})
                    form = BecaTipoConfiguracionForm(initial=model_to_dict(x))
                    form.editar(x)
                    data['form'] = form
                    return render(request, "adm_becas/editbecatipoconfiguracion.html", data)
                except Exception as ex:
                    pass

            elif action == 'reporteria':
                try:
                    data['title'] = u'Reporterías de becas'
                    search = None
                    ids = None
                    data['becaasignacion'] = BecaAsignacion.objects.filter(status=True)
                    return render(request, "adm_becas/reporteria.html", data)
                except Exception as ex:
                    pass

            elif action == 'historialbeca':
                try:
                    data['inscripcion'] = inscripcion = Inscripcion.objects.get(pk=request.GET['id'])
                    becas = BecaAsignacion.objects.filter(status=True, solicitud__inscripcion=inscripcion).order_by(
                        "id")
                    data['becas'] = becas
                    return render(request, "adm_becas/historialbeca.html", data)
                except Exception as ex:
                    pass

            elif action == 'calculobeca':
                try:
                    data['title'] = u'Calculo de becas'
                    data['info'] = None
                    nv = 0.1
                    try:
                        li = float(request.GET['li']) if 'li' in request.GET and request.GET['li'] else 1.0
                        lf = float(request.GET['lf']) if 'lf' in request.GET and request.GET['lf'] else 2.0
                        malla_id = int(request.GET['m']) if 'm' in request.GET and request.GET['m'] else 0
                    except:
                        data['info'] = 'Ocurrio un error en el calculo por motivo de parametros'
                    data['li'] = li
                    data['lf'] = lf
                    data['nv'] = nv
                    malla_ids = Materia.objects.values_list('asignaturamalla__malla_id').filter(nivel__periodo=periodo)
                    data['mallas'] = mallas = Malla.objects.filter(pk__in=malla_ids,
                                                                   carrera__inscripcion__isnull=False,
                                                                   carrera__inscripcion__inscripcionnivel__nivel__orden__gt=1,
                                                                   carrera__coordinacion__in=[1, 2, 3, 4, 5]).exclude(pk__in=[22, 353]).distinct()
                    malla = None
                    if mallas.filter(pk=malla_id).exists():
                        data['malla'] = malla = mallas.filter(pk=malla_id).first()
                    elif mallas.exists():
                        data['malla'] = mallas.first()
                    else:
                        data['malla'] = malla = None
                    aData = []
                    promedio = 0.0
                    desviacion = 0.0

                    if malla:
                        cursor = connection.cursor()
                        sqlnum = """select num_estudiantes_x_carrera_x_periodo_all (%s, 1, %s) """ % (
                            periodo.id, malla.id)
                        cursor.execute(sqlnum)
                        data['num_estudiantes_carrera'] = num_estudiantes_carrera = cursor.fetchone()[0]
                        # print(num_estudiantes_carrera)
                        cursor = connection.cursor()
                        sqlpromedio = """select promedio_carrera_x_periodo_all (%s, 1, %s) """ % (periodo.id, malla.id)
                        cursor.execute(sqlpromedio)
                        data['promedio'] = promedio = float(cursor.fetchone()[0])
                        # print(promedio)
                        sqldesviacion = """select stddev_promedio_carrera_x_periodo_all (%s, 1, %s) """ % (
                            periodo.id, malla.id)
                        cursor.execute(sqldesviacion)
                        data['desviacion'] = desviacion = float(cursor.fetchone()[0])
                        # print(desviacion)

                        while li <= lf + 0.1:
                            v1 = null_to_numeric(li * desviacion, 2)
                            v2 = null_to_numeric(((li * desviacion) + promedio), 2)
                            sqlnum_estudiantes = """select num_estudiantes_x_carrera_x_periodo_promedio_all (%s, 1, %s, %s) """ % (
                                periodo.id, malla.id, v2)
                            cursor.execute(sqlnum_estudiantes)
                            v3 = cursor.fetchone()[0]
                            # print(v3)
                            try:
                                v4 = null_to_numeric(((v3 / num_estudiantes_carrera) * 100), 2)
                            except ZeroDivisionError:
                                v4 = 0

                            aData.append({"num": null_to_numeric(li, 2),
                                          "v1": v1,
                                          "v2": v2,
                                          "v3": v3,
                                          "v4": v4})
                            li += nv
                            data['aData'] = aData
                    return render(request, "adm_becas/calculobeca.html", data)
                except Exception as ex:
                    pass

            elif action == 'reporte_estudiantes_primer_nivel':
                try:
                    reporte = Reporte.objects.get(pk=request.GET['reporte'])
                    tipo = 'xlsx'
                    paRequest = {
                        'periodo_id': periodo.id,
                    }
                    d = run_report_v1(reporte=reporte, tipo=tipo, paRequest=paRequest, request=request)
                    if not d['isSuccess']:
                        raise NameError(d['mensaje'])
                    else:
                        return ok_json({"r": d['mensaje'], 'reportfile': d['data']['reportfile']})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return bad_json(mensaje="Error, al generar el reporte. %s" % ex.__str__())

            elif action == 'listadetallerequisitosbecas':
                try:
                    data['title'] = u'Gestión de Requisitos'
                    search = None
                    ids = None
                    # periodosbeca = BecaSolicitud.objects.filter(status=True).values_list('periodo', flat=True).distinct()
                    # periodos = Periodo.objects.filter(status=True, id__in=periodosbeca)
                    detallesrequisitos = DetalleRequisitoBeca.objects.filter(status=True).order_by('numero')
                    if 's' in request.GET:
                        search = request.GET['s']
                        if search.isdigit():
                            detallesrequisitos = detallesrequisitos.filter(pk=search)
                        else:
                            detallesrequisitos = detallesrequisitos.filter(Q(periodo__nombre__icontains=search))
                    if 'id' in request.GET:
                        ids = request.GET['id']
                        detallesrequisitos = detallesrequisitos.filter(id=ids)
                    # vigente = 0
                    # if 'vigente' in request.GET and int(request.GET['vigente']) > 0:
                    #     vigente = int(request.GET['vigente'])
                    #     periodos = periodos.filter(vigente=int(request.GET['vigente']) == 1)
                    paging = MiPaginador(detallesrequisitos, 25)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['page'] = page
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['detallesrequisitos'] = page.object_list
                    #data['vigente'] = vigente
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    return render(request, "adm_becas/listadetallerequisitosbecas.html", data)
                except Exception as ex:
                    pass

            elif action == 'listarequisitosbecas':
                try:
                    data['title'] = u'Gestión de Requisitos Generales'
                    search = None
                    ids = None
                    # periodosbeca = BecaSolicitud.objects.filter(status=True).values_list('periodo', flat=True).distinct()
                    # periodos = Periodo.objects.filter(status=True, id__in=periodosbeca)
                    requisitosbecas = RequisitoBeca.objects.filter(status=True)
                    if 's' in request.GET:
                        search = request.GET['s']
                        if search.isdigit():
                            requisitosbecas = requisitosbecas.filter(pk=search)
                        else:
                            requisitosbecas = requisitosbecas.filter(Q(nombre__icontains=search))
                    if 'id' in request.GET:
                        ids = request.GET['id']
                        requisitosbecas = requisitosbecas.filter(id=ids)
                    # vigente = 0
                    # if 'vigente' in request.GET and int(request.GET['vigente']) > 0:
                    #     vigente = int(request.GET['vigente'])
                    #     periodos = periodos.filter(vigente=int(request.GET['vigente']) == 1)
                    paging = MiPaginador(requisitosbecas, 25)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['page'] = page
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['requisitosbecas'] = page.object_list
                    #data['vigente'] = vigente
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    return render(request, "adm_becas/listarequisitosbecas.html", data)
                except Exception as ex:
                    pass

            elif action == 'adddetallerequisitobeca':
                try:
                    data['title'] = u'Adicionar Requisito'
                    form = DetalleRequisitoBecaForm()
                    userlog = request.user
                    if userlog.is_superuser:
                        listado = [(chfun[0], f'{chfun[1][0]} \n({chfun[1][1]})') for chfun in FUNCIONES_REQUISITOSBECAS_EJECUTAR]
                        listado.insert(0, ('', '---------'))
                        form.fields['funcionejecutar'].choices = listado
                    data['form'] = form
                    data['action'] = action
                    template = get_template("adm_becas/modal/formdetallerequisitobeca.html")
                    return JsonResponse({"result": True, 'data': template.render(data), 'title': data['title']})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": False, "mensaje": u"Error al adicionar los datos."})

            elif action == 'editdetallerequisitobeca':
                try:
                    data['title'] = u'Editar Requisito'
                    data['id'] = id = int(encrypt(request.GET['id']))
                    data['registro'] = registro = DetalleRequisitoBeca.objects.get(pk=id)
                    form = DetalleRequisitoBecaForm()
                    userlog = request.user
                    if userlog.is_superuser:
                        listado = [(chfun[0], f'{chfun[1][0]} \n({chfun[1][1]})') for chfun in FUNCIONES_REQUISITOSBECAS_EJECUTAR]
                        listado.insert(0, ('', '---------'))
                        form.fields['funcionejecutar'].choices = listado

                    form.initial = model_to_dict(registro)
                    data['form'] = form
                    data['action'] = action

                    template = get_template("adm_becas/modal/formdetallerequisitobeca.html")
                    return JsonResponse({"result": True, 'data': template.render(data), 'title': data['title']})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": False, "mensaje": u"Error al adicionar los datos."})

            elif action == 'showBecaTipoConfiguracion':
                try:
                    data['title'] = u'Detalle Beca Tipo Configuración'
                    data['id'] = id = int(encrypt(request.GET['id']))
                    data['registro'] = registro = BecaTipoConfiguracion.objects.get(pk=id)
                    data['action'] = action
                    data['user'] = request.user
                    template = get_template("adm_becas/modal/showbecatipoconfiguracion.html")
                    return JsonResponse({"result": True, 'data': template.render(data), 'title': data['title']})
                except Exception as ex:
                    msg = str(ex)
                    transaction.set_rollback(True)
                    return JsonResponse({"result": False, "mensaje": u"Error: %s"%msg})

            elif action == 'showrequisitospreinscripcion':
                try:
                    data['title'] = u'Requisito necesarios para obtener beca'
                    data['id'] = id = int(encrypt(request.GET['id']))
                    data['inscripcion'] = inscripcion = Inscripcion.objects.get(id=id)
                    data['persona'] = inscripcion.persona
                    data['registro'] = registro = PreInscripcionBeca.objects.filter(inscripcion_id=id, periodo=periodosesion).first()
                    data['action'] = action
                    data['user'] = usuario = request.user
                    data['puede_cambiar_estado_requisito'] = usuario.has_perm('sga.puede_cambiarestado_requisitos_becas')
                    data['formarchivorequisito'] = SubirArchivoRequisitoBecaForm()
                    template = get_template("adm_becas/modal/showrequisitospreinscripcion.html")
                    return JsonResponse({"result": True, 'data': template.render(data), 'title': data['title']})
                except Exception as ex:
                    msg = str(ex)
                    transaction.set_rollback(True)
                    return JsonResponse({"result": False, "mensaje": u"Error: %s"%msg})

            elif action == 'cambiarestadorequisitobeca':
                try:
                    data['title'] = u'Requisito necesarios para obtener beca'
                    data['id'] = id = int(encrypt(request.GET['id']))
                    data['registro'] = registro = PreInscripcionBecaRequisito.objects.get(id=id)
                    data['action'] = action
                    data['cumplerequisito'] = request.GET['cumplerequisito']
                    data['user'] = usuario = request.user
                    data['puede_cambiar_estado'] = usuario.has_perm('sga.puede_cambiarestado_requisitos_becas')
                    template = get_template("adm_becas/modal/formcambiarestadorequisitobeca.html")
                    return JsonResponse({"result": True, 'data': template.render(data), 'title': data['title']})
                except Exception as ex:
                    msg = str(ex)
                    transaction.set_rollback(True)
                    return JsonResponse({"result": False, "mensaje": u"Error: %s"%msg})

            elif action == 'importar_archivopreinscriptos_becas':
                try:
                    data['id'] = id = int(encrypt(request.GET['id']))
                    data['registro'] = registro = BecaTipo.objects.get(pk=id)
                    data['title'] = u'Importar archivo preinscritos  de tipo beca: %s'%registro
                    data['form'] = ImportarPreinscritosBecaForm()
                    data['action'] = action
                    data['user'] = request.user
                    template = get_template("adm_becas/modal/formimportararchivo.html")
                    return JsonResponse({"result": True, 'data': template.render(data), 'title': data['title']})
                except Exception as ex:
                    msg = str(ex)
                    transaction.set_rollback(True)
                    return JsonResponse({"result": False, "mensaje": u"Error: %s"%msg})

            elif action == 'importar_archivoexcelbecarios':
                try:
                    data['id'] = id = int(encrypt(request.GET['id']))
                    data['registro'] = registro = BecaTipo.objects.get(pk=id)
                    data['title'] = u'Importar archivo becados %s'%registro
                    data['form'] = ImportarArchivoBecariosForm()
                    data['action'] = action
                    data['user'] = request.user
                    template = get_template("adm_becas/modal/formimportararchivobecado.html")
                    return JsonResponse({"result": True, 'data': template.render(data), 'title': data['title']})
                except Exception as ex:
                    msg = str(ex)
                    transaction.set_rollback(True)
                    return JsonResponse({"result": False, "mensaje": u"Error: %s"%msg})

            elif action == 'listadocumentosbecas':
                try:
                    data['title'] = u'Gestión de documentos becas'
                    search = None
                    ids = None
                    # periodosbeca = BecaSolicitud.objects.filter(status=True).values_list('periodo', flat=True).distinct()
                    # periodos = Periodo.objects.filter(status=True, id__in=periodosbeca)
                    detallesrequisitos = DocumentoBecaTipoConfiguracion.objects.filter(status=True).order_by('numero')
                    if 's' in request.GET:
                        search = request.GET['s']
                        if search.isdigit():
                            detallesrequisitos = detallesrequisitos.filter(pk=search)
                        else:
                            detallesrequisitos = detallesrequisitos.filter(Q(periodo__nombre__icontains=search))
                    if 'id' in request.GET:
                        ids = request.GET['id']
                        detallesrequisitos = detallesrequisitos.filter(id=ids)
                    # vigente = 0
                    # if 'vigente' in request.GET and int(request.GET['vigente']) > 0:
                    #     vigente = int(request.GET['vigente'])
                    #     periodos = periodos.filter(vigente=int(request.GET['vigente']) == 1)
                    paging = MiPaginador(detallesrequisitos, 25)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['page'] = page
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['listado'] = page.object_list
                    #data['vigente'] = vigente
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    return render(request, "adm_becas/listadocumentosbecas.html", data)
                except Exception as ex:
                    pass

            elif action == 'adddocumentobecas':
                try:
                    data['title'] = u'Adicionar Requisito'
                    form = DocumentoBecaTipoConfiguracionForm()
                    userlog = request.user
                    if userlog.is_superuser:
                        listado = [(chfun[0], f'{chfun[1][0]} \n({chfun[1][1]})') for chfun in FUNCIONES_DOCUMENTOSBECAS_EJECUTAR]
                        listado.insert(0, ('', '---------'))
                        form.fields['funcionejecutar'].choices = listado
                    data['form'] = form
                    data['action'] = action
                    template = get_template("adm_becas/modal/formdetallerequisitobeca.html")
                    return JsonResponse({"result": True, 'data': template.render(data), 'title': data['title']})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": False, "mensaje": u"Error al adicionar los datos."})

            elif action == 'editdocumentobecas':
                try:
                    data['title'] = u'Editar Requisito'
                    data['id'] = id = int(encrypt(request.GET['id']))
                    data['registro'] = registro = DocumentoBecaTipoConfiguracion.objects.get(pk=id)
                    form = DetalleRequisitoBecaForm()
                    userlog = request.user
                    if userlog.is_superuser:
                        listado = [(chfun[0], f'{chfun[1][0]} \n({chfun[1][1]})') for chfun in FUNCIONES_DOCUMENTOSBECAS_EJECUTAR]
                        listado.insert(0, ('', '---------'))
                        form.fields['funcionejecutar'].choices = listado

                    form.initial = model_to_dict(registro)
                    data['form'] = form
                    data['action'] = action

                    template = get_template("adm_becas/modal/formdetallerequisitobeca.html")
                    return JsonResponse({"result": True, 'data': template.render(data), 'title': data['title']})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": False, "mensaje": u"Error al adicionar los datos."})
                
            elif action == 'reportes_documentacion_becas':
                try:
                    data['title'] = u'Editar Requisito'
                    __author__ = 'Unemi'
                    style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
                    style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
                    style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
                    title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                    style1 = easyxf(num_format_str='D-MMM-YY')
                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('listado_documentacion_becados')
                    ws.write_merge(0, 0, 0, 6, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = f'attachment; filename=reporte_docuementacion_estados{periodo.nombre}' + random.randint(1, 10000).__str__() + '.xls'
                    columns = [
                        (u"ESTUDIANTE", 15000),
                        (u"TIPO DOCUMENTO", 6000),
                        (u"DOCUMENTO", 6000),
                        (u"TIPO BECA", 10000),
                        (u"ESTADO CEDULA", 9000),
                        (u"ESTADO CUENTA BANCARIA", 9000),
                        (u"ESTADO ARCHIVO ENTRENA", 9000),
                        (u"ESTADO ARCHIVO EVENTO", 6000),
                        (u"ESTADO ARCHIVO RAZA", 9000),
                        (u"ESTADO ARCHIVO DISCAPACIDAD", 9000),
                        (u"REQUIERE SERVICIO BASICO", 6000),
                        (u"ESTADO ARCHIVO SERVICIO BASICO", 9000),
                        (u"PAGO SOLICITADO", 9000),
                    ]
                    row_num = 3
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], font_style)
                        ws.col(col_num).width = columns[col_num][1]
                    cursor = connections['sga_select'].cursor()
                    date_format = xlwt.XFStyle()
                    date_format.num_format_str = 'yyyy/mm/dd'
                    sql = f"""
                            SELECT 
                                distinct
                                (persona."apellido1" || ' ' || persona."apellido2"|| ' ' ||persona."nombres" ) AS estudiante,
                                CASE  
                                    WHEN persona.cedula IS NOT NULL OR persona.cedula !=''   THEN 'CEDULA'
                                    WHEN persona.pasaporte IS NOT NULL OR persona.pasaporte !='' THEN 'PASAPORTE'
                                    WHEN persona.ruc IS NOT NULL OR persona.ruc !='' THEN 'RUC'
                                    ELSE ''
                                END tipo_documento,
                                CASE  
                                    WHEN persona.cedula IS NOT NULL OR persona.cedula !=''   THEN persona.cedula
                                    WHEN persona.pasaporte IS NOT NULL OR persona.pasaporte !='' THEN persona.pasaporte
                                    WHEN persona.ruc IS NOT NULL OR persona.ruc !='' THEN persona.ruc
                                    ELSE ''
                                END documento,
                                becatipo.nombre tipo,
                                CASE  docupersonal.estadocedula 
                                    WHEN 1 THEN 'CARGADO'
                                    WHEN 2 THEN 'VALIDADO'
                                    WHEN 3 THEN 'RECHAZADO'
                                    --WHEN 4 THEN 'NO CARGADO'
                                    WHEN 5 THEN 'REVISIÓN'
                                    WHEN 6 THEN 'RECHAZADO IO'
                                    ELSE 'NO CARGADO'
                                END estado_cedula,
                                CASE cuentapersona.estadorevision
                                    WHEN 1 THEN 'CARGADO'
                                    WHEN 2 THEN 'VALIDADO'
                                    WHEN 3 THEN 'RECHAZADO'
                                    WHEN 5 THEN 'REVISIÓN'
                                    WHEN 6 THEN 'RECHAZADO IO'
                                    ELSE 'NO CARGADO'
                                END estado_cuentabancaria,
                                CASE deportista.estadoarchivoentrena
                                    WHEN 1 THEN 'CARGADO'
                                    WHEN 2 THEN 'VALIDADO'
                                    WHEN 3 THEN 'RECHAZADO'
                                    WHEN 5 THEN 'REVISIÓN'
                                    WHEN 6 THEN 'RECHAZADO IO'
                                    ELSE 'NO CARGADO'
                                END estado_archivoentrena,
                                CASE deportista.estadoarchivoevento
                                    WHEN 1 THEN 'CARGADO'
                                    WHEN 2 THEN 'VALIDADO'
                                    WHEN 3 THEN 'RECHAZADO'
                                    WHEN 5 THEN 'REVISIÓN'
                                    WHEN 6 THEN 'RECHAZADO IO'
                                    ELSE 'NO CARGADO'
                                END estado_archivoevento,
                                CASE perfil.estadoarchivoraza
                                    WHEN 1 THEN 'CARGADO'
                                    WHEN 2 THEN 'VALIDADO'
                                    WHEN 3 THEN 'RECHAZADO'
                                    WHEN 5 THEN 'REVISIÓN'
                                    WHEN 6 THEN 'RECHAZADO IO'
                                    ELSE 'NO CARGADO'
                                END estado_raza,
                                CASE perfil.estadoarchivodiscapacidad
                                    WHEN 1 THEN 'CARGADO'
                                    WHEN 2 THEN 'VALIDADO'
                                    WHEN 3 THEN 'RECHAZADO'
                                    WHEN 5 THEN 'REVISIÓN'
                                    WHEN 6 THEN 'RECHAZADO IO'
                                    ELSE 'NO CARGADO'
                                END estado_archivodiscapacidad,
                                CASE 
                                    WHEN (not EXISTS(SELECT * FROM sga_migrantepersona WHERE persona_id=persona.id) AND paisnacimiento_id=1 AND pais_id!=1) THEN 'SI' 
                                    ELSE 'NO'
                                END requiere_serviciobasico,
                                CASE docupersonal.estadoserviciosbasico
                                    WHEN 1 THEN 'CARGADO'
                                    WHEN 2 THEN 'VALIDADO'
                                    WHEN 3 THEN 'RECHAZADO'
                                    WHEN 5 THEN 'REVISIÓN'
                                    WHEN 6 THEN 'RECHAZADO IO'
                                    ELSE 'NO CARGADO'
                                END estado_serviciobasico,
                                CASE WHEN pagobeca.id IS NOT NULL THEN 'SI' ELSE 'NO' END pagosolicitado
                            FROM sga_becasolicitud becasolicitud
                                INNER JOIN sga_becatipo becatipo ON becatipo.id=becasolicitud.becatipo_id
                                INNER JOIN sga_becaasignacion becaasignacion ON becaasignacion.solicitud_id=becasolicitud.id
                                INNER JOIN sga_inscripcion inscripcion ON inscripcion.id=becasolicitud.inscripcion_id
                                INNER JOIN sga_persona persona ON persona.id=inscripcion.persona_id
                                LEFT JOIN sga_personadocumentopersonal docupersonal ON docupersonal.persona_id=persona.id and docupersonal.status OR docupersonal.cedula IS NULL
                                LEFT JOIN sga_cuentabancariapersona cuentapersona ON cuentapersona.persona_id=persona.id and cuentapersona.status AND cuentapersona.activapago
                                LEFT JOIN sga_perfilinscripcion perfil ON perfil.persona_id= persona.id
                                LEFT JOIN sga_deportistapersona deportista ON deportista.persona_id=persona.id
                                LEFT JOIN sga_solicitudpagobecadetalle pagobeca on pagobeca.asignacion_id=becaasignacion.id
                            WHERE becasolicitud.becaaceptada=2
                            AND becasolicitud.periodo_id={periodo.id}
                    """
                    cursor.execute(sql)
                    results = cursor.fetchall()
                    row_num = 4
                    for r in results:
                        estudiante = r[0]
                        tipo_documento = r[1]
                        documento = r[2]
                        tipo_beca = r[3]
                        ectado_cedula = r[4]
                        ectado_cuentabancaria = r[5]
                        estado_archivoentrena = r[6]
                        estado_archivoevento = r[7]
                        estado_raza = r[8]
                        estado_archivodiscapacidad = r[9]
                        requiere_serviciobasico = r[10]
                        estado_serviciobasico = r[11]
                        pagosolicitado = r[12]
                        ws.write(row_num, 0, estudiante, font_style2)
                        ws.write(row_num, 1, tipo_documento, font_style2)
                        ws.write(row_num, 2, documento, font_style2)
                        ws.write(row_num, 3, tipo_beca, font_style2)
                        ws.write(row_num, 4, ectado_cedula, font_style2)
                        ws.write(row_num, 5, ectado_cuentabancaria, font_style2)
                        ws.write(row_num, 6, estado_archivoentrena, font_style2)
                        ws.write(row_num, 7, estado_archivoevento, font_style2)
                        ws.write(row_num, 8, estado_raza, font_style2)
                        ws.write(row_num, 9, estado_archivodiscapacidad, font_style2)
                        ws.write(row_num, 10, requiere_serviciobasico, font_style2)
                        ws.write(row_num, 11, estado_serviciobasico, font_style2)
                        ws.write(row_num, 12, pagosolicitado, font_style2)
                        row_num += 1
                    wb.save(response)
                    connection.close()
                    return response
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": False, "mensaje": u"Error al adicionar los datos."})

            elif action == 'loadFormValidacionMasivoDocumentos':
                try:
                    typeForm = request.GET['typeForm'] if 'typeForm' in request.GET and request.GET['typeForm'] and str(request.GET['typeForm']) in ['new', 'edit', 'view'] else None
                    if typeForm is None:
                        raise NameError(u"No se encontro el tipo de formulario")
                    form = FiltroBecadoValidacionForm()#{'periodopuesto': ultimoperiodo}
                    # form.editar()
                    # data['personaplan'] = PersonaPlanTh.objects.get(pk=int(request.GET['id']))
                    data['form'] = form
                    template = get_template("adm_becas/modal/formvalidaciondocumentosmasivo.html")
                    return JsonResponse({"result": True, 'data': template.render(data, request=request)})
                except Exception as ex:
                    return JsonResponse({"result": False, 'mensaje': ex.__str__()})

            elif action == 'listadopredatabecados':
                try:
                    data['title'] = 'Listado de predata Becados'
                    data['periodo_beca'] = periodo_beca = periodo.becaperiodo_set.filter(status=True).first()
                    data['becatipo_select'] = 0
                    if periodo_beca is not None:
                        data['becatipo_configuraciones'] = periodo_beca.becatipoconfiguracion_set.filter(status=True)
                    return render(request, "adm_becas/viewpredatabecados.html", data)
                except Exception as ex:
                    pass

            elif action == 'adicionarprebecado':
                try:
                    data['title'] = u'Adicionar Precandidato'
                    form = AdicionarPreBecadoForm()
                    form.fields['persona'].queryset = Persona.objects.none()
                    data['form'] = form
                    data['action'] = action
                    return render(request, "adm_becas/addprebecado.html", data)
                except Exception as ex:
                    HttpResponseRedirect(f"{request.path}?info={ex.__str__()}")

            elif action == 'buscarpersonas':
                try:
                    filtro = (Q(inscripcion__status=True, inscripcion__activo=True,
                              inscripcion__matricula__nivel__periodo=periodo,
                              inscripcion__matricula__status=True,
                              inscripcion__matricula__retiradomatricula=False,
                              inscripcion__coordinacion__id__in=[1, 2, 3, 4, 5],
                              perfilusuario__visible=True, perfilusuario__status=True))
                    resp = filtro_persona_select_v2(request, [], filtro)
                    return HttpResponse(json.dumps({'status': True, 'results': resp}))
                except Exception as ex:
                    pass

            elif action == 'rpt_preinscriptos_becas':
                try:
                    tipoarchivo = request.GET.get('tpr')
                    if tipoarchivo == '2':
                        filtros = 'AND prb.seleccionado' if request.GET.get('bsel', '') == '1' else ''
                        sql = f"""
                                SET statement_timeout = 3600000;
                                SELECT 
                                        prb.id "Codigo Preinscripcion",
                                        i.id  "Codigo Inscripcion",  
                                        co.nombre "Coordinacion",
                                        ca.nombre "Carrera",
                                        moda.nombre "Modalidad",
                                        (p.nombres || ' ' || p.apellido1 || ' ' || p.apellido2) "Estudiante",
                                        sex.nombre "Sexo",
                                        p.nacimiento "Fecha Nacimiento",
                                        pa_nac.nombre "Pais Nacimiento",
                                        prov_nac.nombre "Provincia Nacimiento",
                                        can_nac.nombre "Canton Nacimiento",
                                        parr_nac.nombre "Parroquia Nacimiento",
                                        p.cedula "Cedula",
                                        p.email "Email Personal",
                                        p.emailinst "Email Institucional",
                                        p.telefono "Telefono",
                                        p.telefono_conv "Telefono convencional",
                                        pa_res.nombre "Pais Residencia",
                                        prov_res.nombre "Provincia Residencia",
                                        can_res.nombre "Canton Residencia",
                                        parr_res.nombre "Parroquia Residencia",
                                        p.direccion "Direccion",
                                        p.direccion2 "Direccion2",
                                        bt.nombre "Tipo Beca",
                                        CASE WHEN raz.nombre IS NOT NULL THEN raz.nombre ELSE 'NO APLICA' END "Etnia",
                                        CASE WHEN dis.nombre IS NOT NULL THEN dis.nombre ELSE 'NO APLICA' END "Tipo Discapacidad",
                                        CASE WHEN dis.nombre IS NOT NULL THEN prb.carnetdiscapacidad ELSE 'NO APLICA' END "Carnet Discapacidad",
                                        prb.porcientodiscapacidad "Porcentaje Discapacidad",
                                        CASE WHEN migr.id IS NOT NULL THEN 'SI' ELSE 'NO' END "Es migrante",
                                        CASE WHEN dpt.id IS NOT NULL THEN 'SI' ELSE 'NO' END "Es Deportista",
                                        CASE WHEN artp.id IS NOT NULL THEN 'SI' ELSE 'NO' END "Es artista",
                                        nivm.nombre "Semestre",
                                        prb.promedio "Promedio",
                                        prb.orden "Orden",
                                        CASE WHEN prb.seleccionado THEN 'SELECCIONADO' ELSE 'NO SELECCIONAD' END  "Estado becado"
                                FROM 
                                    sga_preinscripcionbeca prb
                                    INNER JOIN sga_inscripcion i ON i.id=prb.inscripcion_id
                                    INNER JOIN sga_persona p ON p.id=i.persona_id
                                    LEFT JOIN sga_pais pa_nac ON pa_nac.id=p.paisnacimiento_id
                                    LEFT JOIN sga_provincia prov_nac ON prov_nac.id=p.provincianacimiento_id
                                    LEFT JOIN sga_canton can_nac ON can_nac.id=p.cantonnacimiento_id
                                    LEFT JOIN sga_parroquia parr_nac ON parr_nac.id=p.parroquianacimiento_id
                                    LEFT JOIN sga_pais pa_res ON pa_res.id=p.pais_id
                                    LEFT JOIN sga_provincia prov_res ON prov_res.id=p.provincia_id
                                    LEFT JOIN sga_canton can_res ON can_res.id=p.canton_id
                                    LEFT JOIN sga_parroquia parr_res ON parr_res.id=p.parroquia_id
                                    INNER JOIN sga_carrera ca ON ca.id=i.carrera_id
                                    INNER JOIN sga_coordinacion co ON co.id=i.coordinacion_id
                                    INNER JOIN sga_modalidad moda ON moda.id=i.modalidad_id
                                    LEFT JOIN sga_sexo sex ON sex.id=p.sexo_id
                                    INNER JOIN sga_becatipo bt ON bt.id=prb.becatipo_id
                                    INNER JOIN sga_matricula mat ON mat.id=prb.matricula_id
                                    INNER JOIN sga_nivelmalla nivm ON nivm.id=mat.nivelmalla_id
                                    LEFT JOIN sga_raza raz ON raz.id=prb.raza_id
                                    LEFT JOIN sga_discapacidad dis ON dis.id=prb.tipodiscapacidad_id
                                    LEFT JOIN sga_migrantepersona migr ON migr.id=prb.migrante_id
                                    LEFT JOIN sga_deportistapersona dpt ON dpt.id=prb.deportista_id
                                    LEFT JOIN sga_artistapersona artp ON artp.id=prb.artista_id
                            WHERE 
                                prb.periodo_id={periodo.id}
                                AND prb."status"
                                {filtros}
                            ORDER BY "Tipo Beca", "Orden"
                        """
                        return generar_reportes_por_query(sql,
                                                          "LISTADO DE PREDATA DE BECADOS",
                                                          "Listado Preinscritos",
                                                          "reporte_excel_prebecados")
                except Exception as ex:
                    pass

            elif action == 'reportependientescargadoc':
                try:
                    return generar_reporte_pendientes_cargar_requisitos(request, periodo)
                except Exception as ex:
                    messages.error(request, str(ex))

            # if action == 'envio_correo_becados_sincuentabancaria_rechazada':
            #     try:
            #         ids_personas = [134212, 118540, 172421, 171585, 157744, 117275, 109689, 170957, 102521, 119144, 60228, 142633, 109960, 96630,
            #          81321,
            #          134401,
            #          112264,
            #          88610,
            #          140712,
            #          169128,
            #          97421,
            #          75847,
            #          85672,
            #          165618,
            #          166027,
            #          118431,
            #          105897,
            #          130997,
            #          80674,
            #          129576,
            #          164245,
            #          158744,
            #          133732,
            #          165002,
            #          159157,
            #          167030,
            #          164057,
            #          156849,
            #          138011,
            #          81115,
            #          131281,
            #          136195,
            #          102203,
            #          107097,
            #          63700,
            #          63480,
            #          118574,
            #          120242,
            #          162512,
            #          119321,
            #          105268,
            #          155453,
            #          108369,
            #          120112,
            #          97016,
            #          134544,
            #          141290,
            #          166196,
            #          73495,
            #          74394,
            #          120363,
            #          167802,
            #          132189,
            #          169258,
            #          82605,
            #          142690,
            #          134168,
            #          134142,
            #          132958,
            #          133135,
            #          97735,
            #          134324,
            #          66291,
            #          139453,
            #          118496,
            #          131165,
            #          135509,
            #          81688,
            #          134943,
            #          140027,
            #          67134,
            #          118917,
            #          87936,
            #          156206,
            #          66452,
            #          102985,
            #          119554,
            #          130406,
            #          134786,
            #          88135,
            #          165869,
            #          134159,
            #          165437,
            #          171618,
            #          134608,
            #          166762,
            #          136648,
            #          128470,
            #          88123,
            #          116017,
            #          49462,
            #          72427,
            #          138431,
            #          128987,
            #          165350,
            #          66784,
            #          166552,
            #          76247,
            #          135280,
            #          112426,
            #          133812,
            #          138776,
            #          169026,
            #          83570,
            #          84809,
            #          139411,
            #          69263,
            #          73501,
            #          134449,
            #          75861,
            #          136115,
            #          138112,
            #          43953,
            #          138358,
            #          167824,
            #          9564,
            #          166378,
            #          75104,
            #          171611,
            #          134183,
            #          145141,
            #          137777,
            #          118656,
            #          133005,
            #          117825,
            #          141004,
            #          69619,
            #          140403,
            #          134302,
            #          116002,
            #          144972,
            #          130828,
            #          59648,
            #          135183,
            #          137154,
            #          154280,
            #          133825,
            #          167823,
            #          157872,
            #          166628,
            #          134163,
            #          165363,
            #          81709,
            #          94452,
            #          128049,
            #          69051,
            #          67188,
            #          166024,
            #          134124,
            #          113179,
            #          134176,
            #          139460,
            #          131337,
            #          109586,
            #          140739,
            #          90012,
            #          156543,
            #          167717,
            #          98870,
            #          144124,
            #          138033,
            #          120068,
            #          143617,
            #          138422,
            #          116582,
            #          168135,
            #          133804,
            #          156266,
            #          135112,
            #          108940,
            #          166155,
            #          140492,
            #          143785,
            #          134292,
            #          79892,
            #          65113,
            #          156731,
            #          143108,
            #          155350,
            #          118480,
            #          168815,
            #          160349,
            #          167516,
            #          67616,
            #          167829,
            #          76513,
            #          162906,
            #          58440,
            #          138522,
            #          134214,
            #          141599,
            #          165398,
            #          135277,
            #          157985,
            #          74003,
            #          134135,
            #          132412,
            #          167914,
            #          165937,
            #          109228,
            #          116324,
            #          136646,
            #          103758,
            #          129138,
            #          132384,
            #          68150,
            #          74129,
            #          168924,
            #          167483,
            #          74473,
            #          156566,
            #          139207,
            #          74437,
            #          141295,
            #          168941,
            #          169636,
            #          100951,
            #          139014,
            #          153882,
            #          134594,
            #          134011,
            #          157868,
            #          155011,
            #          158724,
            #          130712,
            #          117781,
            #          170656,
            #          134622,
            #          166100,
            #          156712,
            #          97788,
            #          74399,
            #          157708,
            #          132559,
            #          169127,
            #          134397,
            #          108732,
            #          118894,
            #          74823,
            #          167591,
            #          156418,
            #          166200,
            #          132376,
            #          139222,
            #          134165,
            #          116599,
            #          171515,
            #          119064,
            #          140687,
            #          143629,
            #          160734,
            #          168959,
            #          100481,
            #          128552,
            #          157494,
            #          166330,
            #          118382,
            #          54735,
            #          132901,
            #          116078,
            #          140405,
            #          85238,
            #          135247,
            #          139642,
            #          102094,
            #          138128,
            #          160054,
            #          119151,
            #          167589,
            #          128503,
            #          166234,
            #          141202,
            #          155585,
            #          153700,
            #          170777,
            #          11127,
            #          10349,
            #          88378,
            #          140959,
            #          143810,
            #          166909,
            #          154338,
            #          156614,
            #          156231,
            #          130531,
            #          131377,
            #          111641,
            #          36338,
            #          158143,
            #          134334,
            #          156354,
            #          168320,
            #          156421,
            #          129316,
            #          139997,
            #          120084,
            #          163096,
            #          157652,
            #          166144,
            #          156724,
            #          169084,
            #          118030,
            #          127711,
            #          127660,
            #          159936,
            #          144810,
            #          167685,
            #          66891,
            #          157693,
            #          168010,
            #          133243,
            #          167578,
            #          157512,
            #          132924,
            #          128731,
            #          78501,
            #          87055,
            #          140281,
            #          160521,
            #          166033,
            #          7967,
            #          86082,
            #          97933,
            #          131843,
            #          132843,
            #          137306,
            #          73443,
            #          159683,
            #          165975,
            #          167588,
            #          156212,
            #          155172,
            #          112648,
            #          93453,
            #          89092,
            #          46141,
            #          59022,
            #          73957,
            #          50916,
            #          43472,
            #          167927,
            #          134521,
            #          118366,
            #          157634,
            #          138728,
            #          141280,
            #          140081,
            #          135058,
            #          167776,
            #          109445,
            #          167780,
            #          165391,
            #          161482,
            #          88877,
            #          135210,
            #          143519,
            #          139105,
            #          133824,
            #          165786,
            #          58115,
            #          116543,
            #          141271,
            #          132029,
            #          139240,
            #          146821,
            #          118380,
            #          139012,
            #          171994,
            #          166102,
            #          158136,
            #          160725,
            #          131938,
            #          119170,
            #          66600,
            #          72672,
            #          170668,
            #          134563,
            #          141409,
            #          169341,
            #          167600,
            #          54822,
            #          111604,
            #          118399,
            #          131231,
            #          75250,
            #          37126,
            #          165994,
            #          128206,
            #          129442,
            #          167757,
            #          14175,
            #          129321,
            #          132480,
            #          85404,
            #          44991,
            #          137989,
            #          120226,
            #          139655,
            #          147101,
            #          19270,
            #          130975,
            #          167603,
            #          132172,
            #          87853,
            #          84427,
            #          128006,
            #          172052,
            #          168258,
            #          160648,
            #          131444,
            #          67791,
            #          113860,
            #          70367,
            #          167158,
            #          134235,
            #          40800,
            #          167539,
            #          144326,
            #          166789,
            #          135137,
            #          86316,
            #          144455,
            #          72316,
            #          24205,
            #          155735,
            #          156516,
            #          144741,
            #          39878,
            #          138692,
            #          160576,
            #          167373,
            #          87624,
            #          135304,
            #          72819,
            #          135268,
            #          37100,
            #          130220,
            #          142825,
            #          134320,
            #          102189,
            #          42271,
            #          166761,
            #          131222,
            #          131867,
            #          158918,
            #          170622,
            #          165266,
            #          86270,
            #          109328,
            #          103351,
            #          141130,
            #          38191,
            #          62608,
            #          133623,
            #          155782,
            #          160728,
            #          135730,
            #          90011,
            #          118251,
            #          67796,
            #          139481,
            #          155736,
            #          139245,
            #          139660,
            #          160652,
            #          130509,
            #          67901,
            #          160013,
            #          95501,
            #          143817,
            #          165789,
            #          100544,
            #          72707,
            #          167073,
            #          120355,
            #          128686,
            #          135233,
            #          133713,
            #          142648,
            #          65645,
            #          143520,
            #          135111,
            #          160155,
            #          163082,
            #          171277,
            #          166962,
            #          95126,
            #          81023,
            #          116951,
            #          165802,
            #          80938,
            #          55204,
            #          68926,
            #          80709,
            #          75743,
            #          138339,
            #          133746,
            #          132270,
            #          119381,
            #          155948,
            #          130107,
            #          166025,
            #          133408,
            #          167263,
            #          128247,
            #          137916,
            #          162966,
            #          84558,
            #          166026,
            #          73934,
            #          172490,
            #          172330,
            #          57612,
            #          67056,
            #          140870,
            #          129862,
            #          162918,
            #          129481,
            #          131364,
            #          138037,
            #          49534,
            #          134308,
            #          82833,
            #          156071,
            #          157814,
            #          75129,
            #          165187,
            #          131673,
            #          133213,
            #          171314,
            #          132725,
            #          140697,
            #          129339,
            #          162942,
            #          67182,
            #          160473,
            #          74347,
            #          132467,
            #          163036,
            #          102192,
            #          131600,
            #          87093,
            #          117811,
            #          133583,
            #          109877,
            #          167216,
            #          171766,
            #          134274,
            #          165880,
            #          155380,
            #          87562,
            #          133298,
            #          35517,
            #          165259,
            #          115998,
            #          155499,
            #          157841,
            #          139213,
            #          73150,
            #          160720,
            #          133229,
            #          108138,
            #          118566,
            #          132156,
            #          119982,
            #          138110,
            #          115967,
            #          127599,
            #          58552,
            #          134065,
            #          164986,
            #          137403,
            #          142156,
            #          110175,
            #          132821,
            #          153704,
            #          165883,
            #          89270,
            #          67422,
            #          138421,
            #          67106,
            #          117028,
            #          139473,
            #          154831,
            #          170212,
            #          159374,
            #          128496,
            #          118428,
            #          99944,
            #          135275,
            #          140224,
            #          134254,
            #          73014,
            #          157817,
            #          166933,
            #          140547,
            #          160627,
            #          169626,
            #          90152,
            #          46136,
            #          169414,
            #          172382,
            #          139298,
            #          157757,
            #          102867,
            #          136287,
            #          165168,
            #          108076,
            #          65235,
            #          169691,
            #          153952,
            #          80728,
            #          141213,
            #          134246,
            #          131000,
            #          133116,
            #          159967,
            #          143143]
            #         import time
            #         for persona_id in ids_personas:
            #             persona = Persona.objects.get(pk=persona_id)
            #             data = {
            #                 'persona': persona,
            #                  'saludo': 'Estimada' if persona.sexo_id == 1 else 'Estimado',
            #             }
            #             list_email = persona.lista_emails_envio()
            #             plantilla = "emails/notificarcuentabancariabeca.html"
            #             send_html_mail("(Recordatorio) Cuentas bancarias para beca estudiantil",
            #                            plantilla,
            #                            data,
            #                            list_email,
            #                            [],
            #                            cuenta=CUENTAS_CORREOS[0][1]
            #                            )
            #             time.sleep(3)
            #
            #     except Exception as ex:
            #         pass
            return HttpResponseRedirect(request.path)
        else:
            try:
                data['title'] = u'SISTEMA DE BECAS'
                search = None
                ids = None

                asignados = BecaAsignacion.objects.filter(status=True, solicitud__periodo=periodo,
                                                          solicitud__becatipo_id__isnull=False).order_by(
                    '-fecha_creacion', 'solicitud__inscripcion__persona__apellido1',
                    'solicitud__inscripcion__persona__apellido2', 'solicitud__inscripcion__persona__nombres')

                c1 = int(null_to_numeric(asignados.filter(~Q(archivocontrato=''),
                                                          archivocontrato__isnull=False,
                                                          solicitud__inscripcion__persona__personadocumentopersonal__status=True,
                                                          solicitud__inscripcion__persona__personadocumentopersonal__estadocedula=2,
                                                          # solicitud__inscripcion__persona__personadocumentopersonal__estadopapeleta=2,
                                                          # solicitud__inscripcion__persona__personadocumentopersonal__estadocedularepresentantesol=2,
                                                          # solicitud__inscripcion__persona__personadocumentopersonal__estadopapeletarepresentantesol=2,
                                                          numerocontrato__isnull=False, cargadocumento=True, tipo=1,
                                                          estadorevisioncontrato=2).exclude(
                    solicitud__becatipo_id=16).aggregate(total=Count('id'))['total'], 2))

                cx = int(null_to_numeric(asignados.filter(~Q(archivocontrato=''),
                                                          archivocontrato__isnull=False,
                                                          solicitud__inscripcion__persona__personadocumentopersonal__status=True,

                                                          numerocontrato__isnull=False, cargadocumento=True, tipo=1,
                                                          estadorevisioncontrato=2).exclude(
                    solicitud__becatipo_id=16).aggregate(total=Count('id'))['total'], 2))

                c2 = int(null_to_numeric(asignados.filter(~Q(archivocontrato=''),
                                                          archivocontrato__isnull=False,
                                                          solicitud__inscripcion__persona__personadocumentopersonal__status=True,
                                                          solicitud__inscripcion__persona__personadocumentopersonal__estadocedula=2,
                                                          # solicitud__inscripcion__persona__personadocumentopersonal__estadopapeleta=2,
                                                          # solicitud__inscripcion__persona__personadocumentopersonal__estadocedularepresentantesol=2,
                                                          # solicitud__inscripcion__persona__personadocumentopersonal__estadopapeletarepresentantesol=2,
                                                          solicitud__inscripcion__persona__personadocumentopersonal__estadoactagrado=2,
                                                          numerocontrato__isnull=False, cargadocumento=True,
                                                          solicitud__becatipo_id=16, tipo=1,
                                                          estadorevisioncontrato=2).aggregate(total=Count('id'))[
                                             'total'], 2))

                c3 = int(null_to_numeric(asignados.filter(~Q(archivocontrato=''),
                                                          archivocontrato__isnull=False,
                                                          solicitud__inscripcion__persona__personadocumentopersonal__status=True,
                                                          numerocontrato__isnull=False, cargadocumento=False,
                                                          tipo=1,
                                                          estadorevisioncontrato=2).aggregate(total=Count('id'))[
                                             'total'], 2))

                totalcontratovalidado = c1 + c2 + c3

                totalbecarenova = asignados.filter(tipo=2).count()

                if 'id' in request.GET:
                    ids = request.GET['id']
                    asignados = asignados.filter(id=search, status=True, solicitud__periodo=periodo).order_by(
                        '-fecha_creacion', 'solicitud__inscripcion__persona__apellido1',
                        'solicitud__inscripcion__persona__apellido2', 'solicitud__inscripcion__persona__nombres')
                elif 's' in request.GET:
                    search = request.GET['s'].strip()
                    ss = search.split(' ')
                    if len(ss) == 1:
                        asignados = asignados.filter(Q(solicitud__inscripcion__persona__cedula__icontains=search) |
                                                     Q(solicitud__inscripcion__persona__nombres__icontains=search) |
                                                     Q(solicitud__inscripcion__persona__apellido1__icontains=search) |
                                                     Q(solicitud__inscripcion__persona__apellido2__icontains=search) |
                                                     Q(solicitud__becatipo__nombre__icontains=search), status=True,
                                                     solicitud__periodo=periodo).order_by('-fecha_creacion',
                                                                                          'solicitud__inscripcion__persona__apellido1',
                                                                                          'solicitud__inscripcion__persona__apellido2',
                                                                                          'solicitud__inscripcion__persona__nombres')
                    else:
                        asignados = asignados.filter(Q(solicitud__inscripcion__persona__nombres__icontains=ss[0]) |
                                                     Q(solicitud__inscripcion__persona__apellido1__icontains=ss[0]) |
                                                     Q(solicitud__inscripcion__persona__apellido2__icontains=ss[0]) |
                                                     Q(solicitud__becatipo__nombre__icontains=ss[0]) |
                                                     Q(solicitud__inscripcion__persona__nombres__icontains=ss[1]) |
                                                     Q(solicitud__inscripcion__persona__apellido1__icontains=ss[1]) |
                                                     Q(solicitud__inscripcion__persona__apellido2__icontains=ss[1]) |
                                                     Q(solicitud__becatipo__nombre__icontains=ss[1])
                                                     , status=True, solicitud__periodo=periodo).order_by(
                            '-fecha_creacion', 'solicitud__inscripcion__persona__apellido1',
                            'solicitud__inscripcion__persona__apellido2', 'solicitud__inscripcion__persona__nombres')

                btiposelect = 0
                if 'btipo' in request.GET:
                    if int(request.GET['btipo']) > 0:
                        btiposelect = int(request.GET['btipo'])
                        asignados = asignados.filter(solicitud__becatipo__id=int(request.GET['btipo'])).order_by(
                            '-fecha_creacion', 'solicitud__inscripcion__persona__apellido1',
                            'solicitud__inscripcion__persona__apellido2', 'solicitud__inscripcion__persona__nombres')
                tiposelect = 0
                if 'tipo' in request.GET:
                    if int(request.GET['tipo']) > 0:
                        tiposelect = int(request.GET['tipo'])
                        asignados = asignados.filter(tipo=int(request.GET['tipo'])).order_by('-fecha_creacion',
                                                                                             'solicitud__inscripcion__persona__apellido1',
                                                                                             'solicitud__inscripcion__persona__apellido2',
                                                                                             'solicitud__inscripcion__persona__nombres')

                estadodocumento = 0
                if 'estadodocumento' in request.GET:
                    estadodocumento = int(request.GET['estadodocumento'])
                    if estadodocumento > 0:
                        if estadodocumento == 1 or estadodocumento == 3:
                            # Q(solicitud__inscripcion__persona__personadocumentopersonal__estadoserviciosbasico=estadodocumento),
                            asignados1 = asignados.filter(
                                Q(solicitud__inscripcion__persona__personadocumentopersonal__estadocedula=estadodocumento),
                                # |
                                # Q(solicitud__inscripcion__persona__personadocumentopersonal__estadopapeleta=estadodocumento) |
                                # Q(solicitud__inscripcion__persona__personadocumentopersonal__estadocedularepresentantesol=estadodocumento) |
                                # Q(solicitud__inscripcion__persona__personadocumentopersonal__estadopapeletarepresentantesol=estadodocumento),
                                cargadocumento=True).exclude(solicitud__becatipo_id=16).order_by('-fecha_creacion',
                                                                                                 'solicitud__inscripcion__persona__apellido1',
                                                                                                 'solicitud__inscripcion__persona__apellido2',
                                                                                                 'solicitud__inscripcion__persona__nombres')

                            asignados2 = asignados.filter(
                                Q(solicitud__inscripcion__persona__personadocumentopersonal__estadocedula=estadodocumento),
                                # Q(solicitud__inscripcion__persona__personadocumentopersonal__estadopapeleta=estadodocumento) |
                                # Q(solicitud__inscripcion__persona__personadocumentopersonal__estadocedularepresentantesol=estadodocumento) |
                                # Q(solicitud__inscripcion__persona__personadocumentopersonal__estadopapeletarepresentantesol=estadodocumento),
                                cargadocumento=True, solicitud__becatipo_id=16).order_by('-fecha_creacion',
                                                                                         'solicitud__inscripcion__persona__apellido1',
                                                                                         'solicitud__inscripcion__persona__apellido2',
                                                                                         'solicitud__inscripcion__persona__nombres')
                            asignados = asignados1 | asignados2
                        else:
                            asignados1 = asignados.filter(cargadocumento=True,
                                                          solicitud__inscripcion__persona__personadocumentopersonal__estadocedula=estadodocumento,
                                                          # solicitud__inscripcion__persona__personadocumentopersonal__estadopapeleta=estadodocumento,
                                                          # solicitud__inscripcion__persona__personadocumentopersonal__estadocedularepresentantesol=estadodocumento,
                                                          # solicitud__inscripcion__persona__personadocumentopersonal__estadopapeletarepresentantesol=estadodocumento,
                                                          ).exclude(solicitud__becatipo_id=16).order_by(
                                '-fecha_creacion', 'solicitud__inscripcion__persona__apellido1',
                                'solicitud__inscripcion__persona__apellido2',
                                'solicitud__inscripcion__persona__nombres')

                            asignados2 = asignados.filter(cargadocumento=True,
                                                          solicitud__inscripcion__persona__personadocumentopersonal__estadocedula=estadodocumento,
                                                          # solicitud__inscripcion__persona__personadocumentopersonal__estadopapeleta=estadodocumento,
                                                          # solicitud__inscripcion__persona__personadocumentopersonal__estadocedularepresentantesol=estadodocumento,
                                                          # solicitud__inscripcion__persona__personadocumentopersonal__estadopapeletarepresentantesol=estadodocumento,
                                                          solicitud__inscripcion__persona__personadocumentopersonal__estadoactagrado=estadodocumento,
                                                          solicitud__becatipo_id=16).order_by('-fecha_creacion',
                                                                                              'solicitud__inscripcion__persona__apellido1',
                                                                                              'solicitud__inscripcion__persona__apellido2',
                                                                                              'solicitud__inscripcion__persona__nombres')

                            asignados = asignados1 | asignados2

                        asignados = asignados.order_by('-fecha_creacion', 'solicitud__inscripcion__persona__apellido1',
                                                       'solicitud__inscripcion__persona__apellido2',
                                                       'solicitud__inscripcion__persona__nombres')

                        asignados = asignados.exclude(tipo=2)

                #                 estadocontrato = 0
                #                 if 'estadocontrato' in request.GET:
                #                     estadocontrato = int(request.GET['estadocontrato'])
                #                     if estadocontrato > 0:
                #                         asignados = asignados.filter(estadorevisioncontrato=estadocontrato)
                #
                #                 estadocumplimiento = estadocomprobante = 0
                #                 if 'estadocumplimiento' in request.GET:
                #                     estadocumplimiento = int(request.GET['estadocumplimiento'])
                #                     if estadocumplimiento > 0:
                #                         asignados = asignados.filter(becacomprobanterevision__isnull=False,
                #                                                      becacomprobanterevision__status=True,
                #                                                      becacomprobanterevision__estadorevisiondbu=estadocumplimiento)
                #
                #                 if 'estadocomprobante' in request.GET:
                #                     estadocomprobante = int(request.GET['estadocomprobante'])
                #                     if estadocomprobante > 0:
                #                         asignados = asignados.filter(becacomprobanterevision__isnull=False,
                #                                                      becacomprobanterevision__status=True,
                #                                                      becacomprobanterevision__estadorevisionfin=estadocomprobante)

                paging = MiPaginador(asignados, 20)
                p = 1
                try:
                    paginasesion = 1
                    if 'paginador' in request.session:
                        paginasesion = int(request.session['paginador'])
                    if 'page' in request.GET:
                        p = int(request.GET['page'])
                    else:
                        p = paginasesion
                    try:
                        page = paging.page(p)
                    except:
                        p = 1
                    page = paging.page(p)
                except:
                    page = paging.page(p)
                data['periodobeca'] = periodobeca = periodo.becaperiodo_set.filter(status=True).first()
                # data['puede_revisar_validar_documentos'] = puede_revisar_validar_documentos = periodobeca.puede_revisar_validar_documentos()
                request.session['paginador'] = p
                data['paging'] = paging
                data['page'] = page
                data['rangospaging'] = paging.rangos_paginado(p)
                data['asignados'] = page.object_list
                data['search'] = search if search else ""
                data['ids'] = ids if ids else ""
                data['tipos'] = TIPO_NUE_RENO
                data['estados'] = ESTADO_BECA_ASIG
                data['tiposelect'] = tiposelect
                data['btiposelect'] = btiposelect
                data['estadodocumento'] = estadodocumento
                # data['estadocontrato'] = estadocontrato
                # data['estadocumplimiento'] = estadocumplimiento
                # data['estadocomprobante'] = estadocomprobante
                # data['becatipo'] = BecaTipo.objects.filter(status=True).order_by('nombre')
                tiposbecaasignadas = BecaSolicitud.objects.values_list('becatipo_id', flat=True).filter(periodo=periodo,
                                                                                                        status=True,
                                                                                                        becaasignada=2).distinct()
                data['becatipo'] = BecaTipo.objects.filter(status=True, pk__in=tiposbecaasignadas).order_by('nombre')
                data['carreras'] = Carrera.objects.filter(status=True, coordinacion__excluir=False,
                                                          id__in=asignados.values_list(
                                                              'solicitud__inscripcion__carrera__id', flat=True))
                data['periodosbecados'] = Periodo.objects.filter(status=True, id__in=BecaAsignacion.objects.values_list(
                    'solicitud__periodo__id', flat=True))
                data['fechaactual'] = datetime.now().strftime('%d-%m-%Y')

                periodovalida = ""
                if BecaAsignacion.objects.values('solicitud__periodocalifica_id').filter(status=True,
                                                                                         solicitud__periodo=periodo,
                                                                                         solicitud__estado=2).exists():
                    # pc = BecaAsignacion.objects.values('solicitud__periodocalifica_id').filter(solicitud__periodo=periodo, status=True, solicitud__estado=2).order_by('-solicitud__periodocalifica_id').distinct()[0]
                    pc = BecaSolicitud.objects.values('periodocalifica_id').filter(status=True, periodo=periodo,
                                                                                   estado=2).order_by(
                        '-periodocalifica_id').distinct()[0]
                    if Periodo.objects.filter(pk=pc['periodocalifica_id']).exists():
                        periodovalida = Periodo.objects.get(pk=pc['periodocalifica_id'])

                data['periodovalida'] = periodovalida

                data['totalbecas'] = totalbecas = asignados.count()
                data['pagosolicitado'] = totalpagosolicitado = asignados.filter(solicitudpagobecadetalle__isnull=False,
                                                                                status=True).distinct().count()
                data['pagonosolicitado'] = asignados.filter(
                    solicitudpagobecadetalle__isnull=True).distinct().count()  # (totalcontratovalidado + totalbecarenova) - totalpagosolicitado

                data['totalcuentas_noregistra'] = asignados.filter(
                    solicitud__inscripcion__persona__cuentabancariapersona__isnull=True).count()
                data['totalcuentas'] = total = asignados.filter(
                    solicitud__inscripcion__persona__cuentabancariapersona__status=True,
                    solicitud__inscripcion__persona__cuentabancariapersona__activapago=True).count()
                data['totalaprobadas'] = aprobadas = asignados.filter(
                    solicitud__inscripcion__persona__cuentabancariapersona__estadorevision=2,
                    solicitud__inscripcion__persona__cuentabancariapersona__activapago=True
                    ).count()
                data['totalrechazadas'] = rechazadas = asignados.filter(
                    solicitud__inscripcion__persona__cuentabancariapersona__estadorevision=3,
                    solicitud__inscripcion__persona__cuentabancariapersona__activapago=True).count()
                data['totalpendiente'] = asignados.filter(
                    solicitud__inscripcion__persona__cuentabancariapersona__isnull=False,
                    solicitud__inscripcion__persona__cuentabancariapersona__estadorevision=1,
                    solicitud__inscripcion__persona__cuentabancariapersona__activapago=True
                    ).count()  # total - (aprobadas + rechazadas)

                totalrenovacion = asignados.filter(tipo=2).count()
                asignados = asignados.exclude(tipo=2)

                nocargadocumento = asignados.filter(cargadocumento=False).count()

                data['totaldoccargada'] = totaldoccargada = asignados.filter(cargadocumento=True,
                                                                             solicitud__inscripcion__persona__personadocumentopersonal__status=True).count()

                totaldocvalida1 = asignados.filter(cargadocumento=True,
                                                   solicitud__inscripcion__persona__personadocumentopersonal__status=True,
                                                   solicitud__inscripcion__persona__personadocumentopersonal__estadocedula=2,
                                                   # solicitud__inscripcion__persona__personadocumentopersonal__estadopapeleta=2,
                                                   # solicitud__inscripcion__persona__personadocumentopersonal__estadocedularepresentantesol=2,
                                                   # solicitud__inscripcion__persona__personadocumentopersonal__estadopapeletarepresentantesol=2
                                                   ).exclude(solicitud__becatipo_id=16).count()

                totaldocvalida2 = asignados.filter(cargadocumento=True,
                                                   solicitud__inscripcion__persona__personadocumentopersonal__status=True,
                                                   solicitud__inscripcion__persona__personadocumentopersonal__estadocedula=2,
                                                   # solicitud__inscripcion__persona__personadocumentopersonal__estadopapeleta=2,
                                                   # solicitud__inscripcion__persona__personadocumentopersonal__estadocedularepresentantesol=2,
                                                   # solicitud__inscripcion__persona__personadocumentopersonal__estadopapeletarepresentantesol=2,
                                                   solicitud__inscripcion__persona__personadocumentopersonal__estadoactagrado=2,
                                                   solicitud__becatipo_id=16).count()

                if periodo.id >= 119:
                    totaldocvalida1 = asignados.filter(  # cargadocumento=True,
                        solicitud__inscripcion__persona__personadocumentopersonal__status=True,
                        solicitud__inscripcion__persona__personadocumentopersonal__estadocedula=2).exclude(
                        solicitud__becatipo_id=16).count()

                    totaldocvalida2 = asignados.filter(cargadocumento=True,
                                                       solicitud__inscripcion__persona__personadocumentopersonal__status=True,
                                                       solicitud__inscripcion__persona__personadocumentopersonal__estadocedula=2,
                                                       solicitud__inscripcion__persona__personadocumentopersonal__estadoactagrado=2,
                                                       solicitud__becatipo_id=16).count()

                data['totaldocvalida'] = totaldocvalida = totaldocvalida1 + totaldocvalida2

                data['totaldocpendcarga'] = (totalbecas - totalrenovacion - nocargadocumento) - totaldoccargada

                totaldocrechazada1 = asignados.filter(
                    Q(solicitud__inscripcion__persona__personadocumentopersonal__estadocedula=3),
                    # Q(solicitud__inscripcion__persona__personadocumentopersonal__estadopapeleta=3) |
                    # Q(solicitud__inscripcion__persona__personadocumentopersonal__estadocedularepresentantesol=3) |
                    # Q(solicitud__inscripcion__persona__personadocumentopersonal__estadopapeletarepresentantesol=3),
                    cargadocumento=True,
                    solicitud__inscripcion__persona__personadocumentopersonal__status=True).exclude(
                    solicitud__becatipo_id=16).count()

                totaldocrechazada2 = asignados.filter(
                    Q(solicitud__inscripcion__persona__personadocumentopersonal__estadocedula=3),  # |
                    # Q(solicitud__inscripcion__persona__personadocumentopersonal__estadopapeleta=3) |
                    # Q(solicitud__inscripcion__persona__personadocumentopersonal__estadocedularepresentantesol=3) |
                    # Q(solicitud__inscripcion__persona__personadocumentopersonal__estadopapeletarepresentantesol=3) |
                    Q(solicitud__inscripcion__persona__personadocumentopersonal__estadoactagrado=3),
                    cargadocumento=True, solicitud__inscripcion__persona__personadocumentopersonal__status=True,
                    solicitud__becatipo_id=16).count()

                data['totaldocrechazada'] = totaldocrechazada = totaldocrechazada1 + totaldocrechazada2
                data['totaldocpendrevision'] = totaldoccargada - (totaldocvalida + totaldocrechazada)

                totalcontgen1 = asignados.filter(numerocontrato__isnull=False, cargadocumento=True).count()
                totalcontgen2 = asignados.filter(numerocontrato__isnull=False, cargadocumento=False).count()

                totalcontpengen2 = asignados.filter(numerocontrato__isnull=True, cargadocumento=False).count()
                totalcontgen = totalcontgen1 + totalcontgen2

                data['totalcontratosgen'] = totalcontgen

                data['totalcontratospendgen'] = (totaldocvalida - totalcontgen1) + totalcontpengen2

                contratos1 = asignados.filter(~Q(archivocontrato=''),
                                              archivocontrato__isnull=False,
                                              solicitud__inscripcion__persona__personadocumentopersonal__status=True,
                                              solicitud__inscripcion__persona__personadocumentopersonal__estadocedula=2,
                                              solicitud__inscripcion__persona__personadocumentopersonal__estadopapeleta=2,
                                              solicitud__inscripcion__persona__personadocumentopersonal__estadocedularepresentantesol=2,
                                              solicitud__inscripcion__persona__personadocumentopersonal__estadopapeletarepresentantesol=2,
                                              numerocontrato__isnull=False, tipo=1).exclude(solicitud__becatipo_id=16)

                contratos2 = asignados.filter(~Q(archivocontrato=''),
                                              archivocontrato__isnull=False,
                                              solicitud__inscripcion__persona__personadocumentopersonal__status=True,
                                              solicitud__inscripcion__persona__personadocumentopersonal__estadocedula=2,
                                              solicitud__inscripcion__persona__personadocumentopersonal__estadopapeleta=2,
                                              solicitud__inscripcion__persona__personadocumentopersonal__estadocedularepresentantesol=2,
                                              solicitud__inscripcion__persona__personadocumentopersonal__estadopapeletarepresentantesol=2,
                                              solicitud__inscripcion__persona__personadocumentopersonal__estadoactagrado=2,
                                              numerocontrato__isnull=False,
                                              solicitud__becatipo_id=16, tipo=1)
                contratoscargados = contratos1 | contratos2

                data['totalcontratoscarga'] = totalcontcargado = contratoscargados.count()
                data['totalcontratospencarga'] = totalcontgen - totalcontcargado
                data['totalcontratosvalida'] = totalcontratosvalida = contratoscargados.filter(
                    estadorevisioncontrato=2).count()
                data['totalcontratosrechazado'] = contratoscargados.filter(estadorevisioncontrato=3).count()
                data['totalcontratospenrevision'] = contratoscargados.filter(estadorevisioncontrato=1).count()

                beneficiarios = Persona.objects.values('id').filter(cuentabancariapersona__status=True,
                                                                    cuentabancariapersona__archivo__isnull=False,
                                                                    inscripcion__becasolicitud__periodo=periodo,
                                                                    inscripcion__becasolicitud__becaasignacion__status=True,
                                                                    inscripcion__becasolicitud__becaasignacion__cargadocumento=True)

                data['compcar'] = compcar = asignados.filter(becacomprobanterevision__isnull=False,
                                                             becacomprobanterevision__status=True).count()

                data['comppencar'] = comppencar = asignados.filter(becacomprobanterevision__isnull=True,
                                                                   solicitudpagobecadetalle__isnull=False,
                                                                   solicitudpagobecadetalle__pagado=True,
                                                                   solicitudpagobecadetalle__acreditado=True).count()

                data['compvaldbu'] = compvaldbu = asignados.filter(becacomprobanterevision__isnull=False,
                                                                   becacomprobanterevision__status=True,
                                                                   becacomprobanterevision__estadorevisiondbu=2).count()

                data['comprecdbu'] = comprecdbu = asignados.filter(becacomprobanterevision__isnull=False,
                                                                   becacomprobanterevision__status=True,
                                                                   becacomprobanterevision__estadorevisiondbu__in=[3,
                                                                                                                   6]).count()

                data['comppendbu'] = compcar - (compvaldbu + comprecdbu)

                data['compvaltes'] = compvaltes = asignados.filter(becacomprobanterevision__isnull=False,
                                                                   becacomprobanterevision__status=True,
                                                                   becacomprobanterevision__estadorevisionfin=2).count()

                data['comprectes'] = comprectes = asignados.filter(becacomprobanterevision__isnull=False,
                                                                   becacomprobanterevision__status=True,
                                                                   becacomprobanterevision__estadorevisionfin__in=[3,
                                                                                                                   6]).count()

                data['comppentes'] = compvaldbu - (compvaltes + comprectes)

                data['nocargadocumento'] = nocargadocumento + totalrenovacion

                if totalpagosolicitado > 0:
                    data['totalpagados'] = totalpagados = SolicitudPagoBecaDetalle.objects.values("id").filter(
                        status=True, pagado=True, solicitudpago__periodo=periodo).count()
                    data['totalporpagar'] = totalpagosolicitado - totalpagados
                    data['totalacreditados'] = totalacreditado = SolicitudPagoBecaDetalle.objects.values("id").filter(
                        status=True, pagado=True, acreditado=True, solicitudpago__periodo=periodo).count()
                    data['totalporacreditar'] = totalpagados - totalacreditado

                view = "view2.html" if periodo.id >= 177 else "view.html"

                return render(request, f"adm_becas/{view}", data)
            except Exception as ex:
                pass



class GenerateBackground(threading.Thread):
    def __init__(self, request, data, periodoactual, periodoanterior, noti=None):
        self.request = request
        self.data = data
        self.periodoactual = periodoactual
        self.periodoanterior = periodoanterior
        self.noti = noti
        threading.Thread.__init__(self)

    def run(self):
        request, data, periodoactual, periodoanterior = self.request, self.data, self.periodoactual, self.periodoanterior
        if request.method == 'POST':
            action = request.POST['action']
            if action == 'generarprecandidatos':
                return generate_precandidatos_beca(request, data, periodoactual, periodoanterior)
            elif action == 'generarsolicitudesbecas':
                return crear_solicitudes_becas(request, periodoactual, periodoanterior, True)
            elif action == 'generarprecandidatosdesviacionestandar':
                return generate_precandidatos_beca_desviacion_standar_todas_carreras(request, data, periodoactual, periodoanterior)
            elif action == 'generarprecandidatosdesviacionestandarnew':
                return generate_precandidatos_beca_desviacion_standar_todas_carrerasnew(request, data, periodoactual, periodoanterior)
            elif action == 'addupdatepreinscripcionbecarequisitos':
                return generar_requisitos_preinscripcionesbecas(request, data, periodoactual, periodoanterior)
            elif action == 'importar_archivopreinscriptos_becas':
                return importar_archivopreinscriptos_becas(request, data, periodoactual, periodoanterior, noti=self.noti)
            elif action == 'importar_archivoexcelbecarios':
                return importar_archivoexcelbecarios(request, data, periodoactual, periodoanterior, noti=self.noti, enviarcorreo=True)
            elif action == 'corregir_actascompromisos_becas':
                return corregir_actascompromisos_becas(request, data, periodoactual, noti=self.noti)
            elif action == 'generate_reporte_pendientes_pago_becas_financiero':
                return generate_reporte_pendientes_pago_becas_financiero(request, data, periodoactual, noti=self.noti)
            elif action == 'enviarcorreopreseleccion':
                return enviarcorreo_precandidatos_beca_todas_carreras(request, data, periodoactual, periodoanterior, self.noti)


def generate_precandidatos_beca(request, data, periodoactual, periodoanterior):
    error = False
    mensaje_ex = None
    with transaction.atomic():
        try:
            becatipos = BecaTipo.objects.filter(status=True, vigente=True)
            ID_MEJOR_PROMEDIO = 17
            ID_DISCAPACIDAD = 19
            ID_DEPORTISTA = 20
            ID_EXTERIOR_MIGRANTE = 22
            ID_ETNIA = 21
            ID_GRUPO_VULNERABLE = 18
            EXCLUDES = []
            """LISTADO DE LOS MEJORES PROMEDIOS POR MALLA"""
            mejores = lista_mejores_promedio_beca(periodoactual=periodoactual, periodoanterior=None)["lista"]
            inscripciones_mejores = Inscripcion.objects.filter(pk__in=mejores).exclude(persona_id__in=PreInscripcionBeca.objects.values("inscripcion__persona_id").filter(periodo=periodoactual))
            actualizar_mejores = PreInscripcionBeca.objects.filter(becatipo=becatipos.filter(pk=ID_MEJOR_PROMEDIO).first(), periodo=periodoactual).exclude(inscripcion__in=inscripciones_mejores)
            if DEBUG:
                actualizar_mejores.delete()
            """COMENTAR CUANDO SE VAYA A PRODUCCIÓN"""
            # inscripciones_mejores = PreInscripcionBeca.objects.filter(becatipo=becatipos.filter(pk=ID_MEJOR_PROMEDIO).first(), periodo=periodo)
            # mejores = list(inscripciones_mejores.values_list('inscripcion_id', flat=True))
            # inscripciones_mejores = Inscripcion.objects.filter(pk__in=mejores)
            """FINAL DEL COMENTADO"""
            for inscripcion in inscripciones_mejores:
                if not PreInscripcionBeca.objects.values("id").filter(inscripcion=inscripcion, periodo=periodoactual).exists():
                    preinscripcion = PreInscripcionBeca(inscripcion=inscripcion,
                                                        promedio=inscripcion.promedio,
                                                        becatipo=becatipos.filter(pk=ID_MEJOR_PROMEDIO).first(),
                                                        periodo=periodoactual,
                                                        fecha=datetime.now().date())
                    preinscripcion.save(request)
            EXCLUDES.extend(mejores)
            """LISTADO DE DISCAPACITADOS EXCLUYENDO A LOS DE MEJORES PROMEDIOS"""
            discapacitados = lista_discapacitado_beca(periodoactual=periodoactual, periodoanterior=periodoanterior, excludes=EXCLUDES)
            inscripciones_discapacitados = Inscripcion.objects.filter(pk__in=discapacitados).exclude(persona_id__in=PreInscripcionBeca.objects.values("inscripcion__persona_id").filter(periodo=periodoactual))
            actualizar_discapacitados = PreInscripcionBeca.objects.filter(becatipo=becatipos.filter(pk=ID_DISCAPACIDAD).first(), periodo=periodoactual).exclude(inscripcion__in=inscripciones_discapacitados)
            if DEBUG:
                actualizar_discapacitados.delete()
            for inscripcion in inscripciones_discapacitados:
                if not PreInscripcionBeca.objects.filter(inscripcion=inscripcion, periodo=periodoactual).exists():
                    preinscripcion = PreInscripcionBeca(inscripcion=inscripcion,
                                                        promedio=inscripcion.promedio,
                                                        becatipo=becatipos.filter(pk=ID_DISCAPACIDAD).first(),
                                                        periodo=periodoactual,
                                                        fecha=datetime.now().date())
                    preinscripcion.save(request)
            EXCLUDES.extend(discapacitados)
            """LISTADO DE DEPORTISTAS EXCLUYENDO A DISCAPACITADOS Y LOS DE MEJORES PROMEDIOS"""
            deportistas = lista_deportista_beca(periodoactual=periodoactual, periodoanterior=periodoanterior, excludes=EXCLUDES)
            inscripciones_deportistas = Inscripcion.objects.filter(pk__in=deportistas).exclude(persona_id__in=PreInscripcionBeca.objects.values("inscripcion__persona_id").filter(periodo=periodoactual))
            actualizar_deportistas = PreInscripcionBeca.objects.filter(becatipo=becatipos.filter(pk=ID_DEPORTISTA).first(), periodo=periodoactual).exclude(inscripcion__in=inscripciones_deportistas)
            if DEBUG:
                actualizar_deportistas.delete()
            for inscripcion in inscripciones_deportistas:
                if not PreInscripcionBeca.objects.filter(inscripcion=inscripcion, periodo=periodoactual).exists():
                    preinscripcion = PreInscripcionBeca(inscripcion=inscripcion,
                                                        promedio=inscripcion.promedio,
                                                        becatipo=becatipos.filter(pk=ID_DEPORTISTA).first(),
                                                        periodo=periodoactual,
                                                        fecha=datetime.now().date())
                    preinscripcion.save(request)
            EXCLUDES.extend(deportistas)
            """LISTADO DE PERSONAS EN EL EXTRANJERO EXCLUYENDO A DISCAPACITADOS, DEPORTISTAS Y LOS DE MEJORES PROMEDIOS"""
            migrantes = lista_migrante_exterior_beca(periodoactual=periodoactual, periodoanterior=periodoanterior, excludes=EXCLUDES)
            inscripciones_migrantes = Inscripcion.objects.filter(pk__in=migrantes).exclude(persona_id__in=PreInscripcionBeca.objects.values("inscripcion__persona_id").filter(periodo=periodoactual))
            actualizar_migrantes = PreInscripcionBeca.objects.filter(becatipo=becatipos.filter(pk=ID_EXTERIOR_MIGRANTE).first(), periodo=periodoactual).exclude(inscripcion__in=inscripciones_migrantes)
            if DEBUG:
                actualizar_migrantes.delete()
            for inscripcion in inscripciones_migrantes:
                if not PreInscripcionBeca.objects.filter(inscripcion=inscripcion, periodo=periodoactual).exists():
                    preinscripcion = PreInscripcionBeca(inscripcion=inscripcion,
                                                        promedio=inscripcion.promedio,
                                                        becatipo=becatipos.filter(pk=ID_EXTERIOR_MIGRANTE).first(),
                                                        periodo=periodoactual,
                                                        fecha=datetime.now().date())
                    preinscripcion.save(request)
            EXCLUDES.extend(migrantes)
            """LISTADO DE PERSONAS ETNIA EXCLUYENDO EXTRANJERO, MIGRANTES, DISCAPACITADOS, DEPORTISTAS Y LOS DE MEJORES PROMEDIOS"""
            etnias = lista_etnia_beca(periodoactual=periodoactual, periodoanterior=periodoanterior, excludes=EXCLUDES)
            inscripciones_etnias = Inscripcion.objects.filter(pk__in=etnias).exclude(persona_id__in=PreInscripcionBeca.objects.values("inscripcion__persona_id").filter(periodo=periodoactual))
            actualizar_etnias = PreInscripcionBeca.objects.filter(becatipo=becatipos.filter(pk=ID_ETNIA).first(), periodo=periodoactual).exclude(inscripcion__in=inscripciones_etnias)
            if DEBUG:
                actualizar_etnias.delete()
            for inscripcion in inscripciones_etnias:
                if not PreInscripcionBeca.objects.filter(inscripcion=inscripcion, periodo=periodoactual).exists():
                    preinscripcion = PreInscripcionBeca(inscripcion=inscripcion,
                                                        promedio=inscripcion.promedio,
                                                        becatipo=becatipos.filter(pk=ID_ETNIA).first(),
                                                        periodo=periodoactual,
                                                        fecha=datetime.now().date())
                    preinscripcion.save(request)
            EXCLUDES.extend(etnias)

            # pruebas = PreInscripcionBeca.objects.filter(periodo=periodo).exclude(becatipo=becatipos.filter(pk=ID_GRUPO_VULNERABLE).first())
            # EXCLUDES.extend(list(pruebas.values_list('inscripcion_id', flat=True).distinct()))
            """LISTADO DE PERSONAS DE GRUPO VULNERABLE EXCLUYENDO ETNIA, EXTRANJERO, MIGRANTES, DISCAPACITADOS, DEPORTISTAS Y LOS DE MEJORES PROMEDIOS"""
            grupo = []
            for grupo_id in [4, 5]:
                grupo.extend(lista_gruposocioeconomico_beca(periodoactual=periodoactual, periodoanterior=periodoanterior, tipogrupo_id=grupo_id, excludes=EXCLUDES, limit=20))
            inscripciones_grupo = Inscripcion.objects.filter(pk__in=grupo).exclude(persona_id__in=PreInscripcionBeca.objects.values("inscripcion__persona_id").filter(periodo=periodoactual))
            actualizar_grupo = PreInscripcionBeca.objects.filter(becatipo=becatipos.filter(pk=ID_GRUPO_VULNERABLE).first(), periodo=periodoactual).exclude(inscripcion__in=inscripciones_grupo)
            if DEBUG:
                actualizar_grupo.delete()
            for inscripcion in inscripciones_grupo:
                if not PreInscripcionBeca.objects.filter(inscripcion=inscripcion, periodo=periodoactual).exists():
                    preinscripcion = PreInscripcionBeca(inscripcion=inscripcion,
                                                        promedio=inscripcion.promedio,
                                                        becatipo=becatipos.filter(pk=ID_GRUPO_VULNERABLE).first(),
                                                        periodo=periodoactual,
                                                        fecha=datetime.now().date())
                    preinscripcion.save(request)
            error = False

        except Exception as ex:
            transaction.set_rollback(True)
            print(ex)
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            error = True
            mensaje_ex = ex.__str__()

    if error:
        notificacion = Notificacion(titulo=f"Error en el proceso de creación de precandidatos de becas",
                                    cuerpo=f"Ocurrio un error al generar precandidatos de becas del periodo académico {periodoactual.nombre}. {mensaje_ex}",
                                    destinatario=request.session['persona'],
                                    url="/adm_becas",
                                    content_type=None,
                                    object_id=None,
                                    prioridad=2,
                                    app_label=request.session['tiposistema'],
                                    fecha_hora_visible=datetime.now() + timedelta(days=2)
                                    )
        notificacion.save(request)
    else:
        notificacion = Notificacion(titulo=f"Proceso finalizado de creación de precandidatos de becas",
                                    cuerpo=f"Se genero correctamente el precandidatos de becas del periodo académico {periodoactual.nombre}",
                                    destinatario=request.session['persona'],
                                    url="/adm_becas",
                                    content_type=None,
                                    object_id=None,
                                    prioridad=2,
                                    app_label=request.session['tiposistema'],
                                    fecha_hora_visible=datetime.now() + timedelta(days=2)
                                    )
        notificacion.save(request)


def crear_solicitudes_becas(request, periodo, periodovalida, enviarcorreo=False):
    error = False
    mensaje_ex = None
    lista_envio_masivo = []
    with transaction.atomic():
        try:
            print("Proceso iniciado...")
            excludes = BecaSolicitud.objects.filter(periodo=periodo).values_list('inscripcion_id', flat=True)
            becatipos = BecaTipo.objects.filter(pk__in=PreInscripcionBeca.objects.values_list('becatipo_id', flat=True).filter(status=True, periodo=periodo).distinct())
            for becatipo in becatipos:
                print("--- Beca tipo: %s" % becatipo.__str__())
                preinscripciones = PreInscripcionBeca.objects.filter(status=True, periodo=periodo, becatipo=becatipo).exclude(inscripcion_id__in=excludes)
                if DEBUG:
                    preinscripciones = preinscripciones.filter(inscripcion_id=104493)
                    # raise NameError("Error provocado")
                for preinscripcion in preinscripciones:
                    persona = preinscripcion.inscripcion.persona
                    if not BecaSolicitud.objects.filter(inscripcion=preinscripcion.inscripcion, periodo=periodo).exists():
                        becado = BecaSolicitud(inscripcion=preinscripcion.inscripcion,
                                               becatipo=becatipo,
                                               periodo=periodo,
                                               periodocalifica=periodovalida,
                                               estado=1,
                                               tiposolicitud=preinscripcion.tipo_renovacion_nueva(periodovalida),
                                               observacion=f'SEMESTRE REGULAR {periodo.nombre}')
                        becado.save(usuario_id=persona.usuario.id)
                    else:
                        becado = BecaSolicitud.objects.filter(inscripcion=preinscripcion.inscripcion, periodo=periodo).first()
                    requisitos = BecaRequisitos.objects.filter(Q(becatipo__isnull=True), becatipo=becatipo, periodo=periodo, status=True, vigente=True).exclude(numero=7).order_by('numero')
                    # print(requisitos)
                    for requisito in requisitos:
                        if requisito.numero == 1 or requisito.numero == 2:
                            if requisito.numero == 1:
                                if persona.paisnacimiento_id == 1:
                                    if not BecaDetalleSolicitud.objects.filter(status=True, solicitud=becado,
                                                                               requisito=requisito, cumple=True,
                                                                               estado=2).exists():
                                        detalle = BecaDetalleSolicitud(solicitud=becado,
                                                                       requisito=requisito,
                                                                       cumple=True,
                                                                       archivo=None,
                                                                       estado=2,
                                                                       observacion='APROBACIÓN AUTOMÁTICA',
                                                                       personaaprueba=None,
                                                                       fechaaprueba=None)
                                        detalle.save(usuario_id=1)
                            else:
                                if persona.paisnacimiento_id != 1:
                                    if not BecaDetalleSolicitud.objects.filter(status=True, solicitud=becado, requisito=requisito, cumple=True, estado=2).exists():
                                        detalle = BecaDetalleSolicitud(solicitud=becado,
                                                                       requisito=requisito,
                                                                       cumple=True,
                                                                       archivo=None,
                                                                       estado=2,
                                                                       observacion='APROBACIÓN AUTOMÁTICA',
                                                                       personaaprueba=None,
                                                                       fechaaprueba=None)
                                        detalle.save(usuario_id=1)
                        else:
                            if not BecaDetalleSolicitud.objects.filter(status=True, solicitud=becado, requisito=requisito, cumple=True, estado=1 if requisito.id in [15, 16] else 2).exists():
                                detalle = BecaDetalleSolicitud(solicitud=becado,
                                                               requisito=requisito,
                                                               cumple=True,
                                                               archivo=None,
                                                               estado=1 if requisito.id in [15, 16] else 2,
                                                               observacion='' if requisito.id in [15, 16] else 'APROBACIÓN AUTOMÁTICA',
                                                               personaaprueba=None,
                                                               fechaaprueba=None)
                                detalle.save(usuario_id=1)

                    # necesidadsolicitud = BecaSolicitudNecesidad(solicitud=beca, necesidad=necesidad)
                    # necesidadsolicitud.save()
                    if not BecaSolicitudRecorrido.objects.filter(solicitud=becado, estado=1).exists():
                        recorrido = BecaSolicitudRecorrido(solicitud=becado,
                                                           observacion="SOLICITUD AUTOMÁTICA",
                                                           estado=1)
                        recorrido.save(usuario_id=1)
                        notificacion = Notificacion(titulo=f"Registro de Solicitud de Beca - {periodo.nombre}",
                                                    cuerpo=f"Registró de solicitud de aplicación a la beca por {becado.becatipo.nombre.upper()} para el periodo académico {periodo.nombre} el {datetime.now().strftime('d-m-Y')} a las { datetime.now().strftime('h:i a') }",
                                                    destinatario=persona,
                                                    url="/alu_becas",
                                                    content_type=None,
                                                    object_id=None,
                                                    prioridad=2,
                                                    app_label=request.session['tiposistema'],
                                                    fecha_hora_visible=datetime.now() + timedelta(days=2)
                                                    )
                        notificacion.save(usuario_id=1)

                        if enviarcorreo:

                            tituloemail = f"Registro de Solicitud de Beca - {periodo.nombre}"
                            data = {'sistema': u'SGA - UNEMI',
                                    'tipo': 'SOL',
                                    'tipobeca': becado.becatipo.nombre.upper(),
                                    'fecha': datetime.now().date(),
                                    'hora': datetime.now().time(),
                                    'periodo': periodo.nombre,
                                    'estudiante': persona,
                                    'autoridad2': '',
                                    't': miinstitucion()
                                    }
                            list_email = persona.lista_emails_envio()
                            if DEBUG:
                                list_email = ['crodriguezn@unemi.edu.ec']
                            lista_envio_masivo.append({'tituloemail': tituloemail,
                                                       'data': data,
                                                       'list_email': list_email,
                                                       'plantilla': "emails/solicitudbecaestudiante.html"})
                            # send_html_mail(tituloemail,
                            #                "emails/solicitudbecaestudiante.html",
                            #                data,
                            #                list_email,
                            #                [],
                            #                cuenta=variable_valor('CUENTAS_CORREOS')[0]
                            #                )
                            # time.sleep(5)
                    # REGISTRO EN ESTADO DE REVISION
                    if not BecaSolicitudRecorrido.objects.filter(solicitud=becado, estado=4).exists():
                        becado.estado = 4
                        becado.save(request)
                        recorrido = BecaSolicitudRecorrido(solicitud=becado,
                                                           observacion="EN REVISION",
                                                           estado=4)
                        recorrido.save(usuario_id=1)

                    if not BecaSolicitudRecorrido.objects.filter(solicitud=becado, estado=2).exists():
                        becado.estado = 2
                        becado.becaaceptada = 1
                        becado.save(request)
                        recorrido = BecaSolicitudRecorrido(solicitud=becado,
                                                           observacion="PENDIENTE DE ACEPTACIÓN O RECHAZO",
                                                           estado=2)
                        recorrido.save(usuario_id=1)

                        notificacion = Notificacion(titulo=f"Solicitud de Beca en Revisión - {periodo.nombre}",
                                                    cuerpo=f"Ssolicitud de aplicación a la beca por {becado.becatipo.nombre.upper()} para el periodo académico {periodo.nombre} ha sido { 'APROBADA' if becado.estado == 2 else 'RECHAZADA' }",
                                                    destinatario=persona,
                                                    url="/alu_becas",
                                                    content_type=None,
                                                    object_id=None,
                                                    prioridad=2,
                                                    app_label=request.session['tiposistema'],
                                                    fecha_hora_visible=datetime.now() + timedelta(days=2)
                                                    )
                        notificacion.save(usuario_id=1)
                        # Envio de e-mail de notificación al solicitante
                        if enviarcorreo:
                            tituloemail = "Solicitud de Beca en Revisión"
                            data = {'sistema': u'SGA - UNEMI',
                                    'fase': 'AR',
                                    'tipobeca': becado.becatipo.nombre.upper(),
                                    'fecha': datetime.now().date(),
                                    'hora': datetime.now().time(),
                                    'saludo': 'Estimada' if becado.inscripcion.persona.sexo_id == 1 else 'Estimado',
                                    'estado': 'APROBADA' if becado.estado == 2 else "RECHAZADA",
                                    'estudiante': becado.inscripcion.persona.nombre_completo_inverso(),
                                    'autoridad2': '',
                                    'observaciones': '',
                                    'periodo': periodo.nombre,
                                    't': miinstitucion()
                                    }
                            list_email = persona.lista_emails_envio()
                            if DEBUG:
                                list_email = ['crodriguezn@unemi.edu.ec']
                            lista_envio_masivo.append({'tituloemail': tituloemail,
                                                       'data': data,
                                                       'list_email': list_email,
                                                       'plantilla': "emails/notificarestadosolicitudbeca.html"})
                            # send_html_mail(tituloemail,
                            #                "emails/notificarestadosolicitudbeca.html",
                            #                data,
                            #                list_email,
                            #                [],
                            #                cuenta=variable_valor('CUENTAS_CORREOS')[0]
                            #                )
                            # time.sleep(5)
                    print("------ Solicitud creadas/aprobadas: ", persona)
            print("Proceso terminado...")
            error = False
        except Exception as ex:
            transaction.set_rollback(True)
            print(ex)
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            error = True
            mensaje_ex = ex.__str__()
            lista_envio_masivo = []

    if error:
        notificacion = Notificacion(titulo=f"Error en el proceso de creación de solicitudes de becas",
                                    cuerpo=f"Ocurrio un error al generar solicitudes de becas del periodo académico {periodo.nombre}. {mensaje_ex}",
                                    destinatario=request.session['persona'],
                                    url="/adm_becas",
                                    content_type=None,
                                    object_id=None,
                                    prioridad=2,
                                    app_label=request.session['tiposistema'],
                                    fecha_hora_visible=datetime.now() + timedelta(days=2)
                                    )
        notificacion.save(usuario_id=1)
    else:
        if enviarcorreo:
            for x in lista_envio_masivo:
                tituloemail = x['tituloemail']
                data = x['data']
                list_email = x['list_email']
                plantilla = x['plantilla']

                send_html_mail(tituloemail,
                               plantilla,
                               data,
                               list_email,
                               [],
                               cuenta=CUENTAS_CORREOS[0][1]
                               )
                time.sleep(5)
        notificacion = Notificacion(titulo=f"Proceso finalizado de creación de solicitudes de becas",
                                    cuerpo=f"Se genero correctamente solicitudes de becas del periodo académico {periodo.nombre}",
                                    destinatario=request.session['persona'],
                                    url="/adm_becas",
                                    content_type=None,
                                    object_id=None,
                                    prioridad=2,
                                    app_label=request.session['tiposistema'],
                                    fecha_hora_visible=datetime.now() + timedelta(days=2)
                                    )
        notificacion.save(usuario_id=1)


def generate_precandidatos_beca_desviacion_standar_todas_carreras(request, data, periodoactual, periodoanterior):
    error = False
    mensaje_ex = None
    with transaction.atomic():
        try:
            becatipos = BecaTipo.objects.filter(status=True, vigente=True)
            ID_PRIMER_NIVEL = 16
            ID_MEJOR_PROMEDIO = 17
            ID_DISCAPACIDAD = 19
            ID_DEPORTISTA = 20
            ID_EXTERIOR_MIGRANTE = 22
            ID_ETNIA = 21
            ID_GRUPO_VULNERABLE = 18
            EXCLUDES = []
            """LISTADO DE INSCRIPCIONES DE PRIMER NIVEL CON RECONOCIMIENTO ACADEMICO"""
            destacadosprimernivel = listado_incripciones_reconocimiento_academico(periodoactual)
            inscripciones_destacadosprimernivel = destacadosprimernivel.exclude(persona_id__in=PreInscripcionBeca.objects.values("inscripcion__persona_id").filter(periodo=periodoactual))
            actualizar_destacadosprimernivel = PreInscripcionBeca.objects.filter(becatipo=becatipos.filter(pk=ID_PRIMER_NIVEL).first(), periodo=periodoactual).exclude(inscripcion__in=inscripciones_destacadosprimernivel)
            if DEBUG:
                actualizar_destacadosprimernivel.delete()

            for inscripcion in inscripciones_destacadosprimernivel:
                if not PreInscripcionBeca.objects.values("id").filter(inscripcion=inscripcion, periodo=periodoactual).exists():
                    preinscripcion = PreInscripcionBeca(inscripcion=inscripcion,
                                                        promedio=inscripcion.promedio,
                                                        becatipo=becatipos.filter(pk=ID_PRIMER_NIVEL).first(),
                                                        periodo=periodoactual,
                                                        fecha=datetime.now().date())
                    preinscripcion.save(request)
            EXCLUDES.extend(destacadosprimernivel)

            """LISTADO DE LOS MEJORES PROMEDIOS POR MALLA"""
            #mejores = lista_mejores_promedio_beca(periodoactual=periodoactual, periodoanterior=None)
            mejores = listado_becados_por_desviacionestandar_todas_carrera(periodoactual)
            inscripciones_mejores = mejores.exclude(persona_id__in=PreInscripcionBeca.objects.values("inscripcion__persona_id").filter(periodo=periodoactual))
            actualizar_mejores = PreInscripcionBeca.objects.filter(becatipo=becatipos.filter(pk=ID_MEJOR_PROMEDIO).first(), periodo=periodoactual).exclude(inscripcion__in=inscripciones_mejores)
            if DEBUG:
                actualizar_mejores.delete()
            """COMENTAR CUANDO SE VAYA A PRODUCCIÓN"""
            # inscripciones_mejores = PreInscripcionBeca.objects.filter(becatipo=becatipos.filter(pk=ID_MEJOR_PROMEDIO).first(), periodo=periodo)
            # mejores = list(inscripciones_mejores.values_list('inscripcion_id', flat=True))
            # inscripciones_mejores = Inscripcion.objects.filter(pk__in=mejores)
            """FINAL DEL COMENTADO"""
            for inscripcion in inscripciones_mejores:
                if not PreInscripcionBeca.objects.values("id").filter(inscripcion=inscripcion, periodo=periodoactual).exists():
                    preinscripcion = PreInscripcionBeca(inscripcion=inscripcion,
                                                        promedio=inscripcion.promedio,
                                                        becatipo=becatipos.filter(pk=ID_MEJOR_PROMEDIO).first(),
                                                        periodo=periodoactual,
                                                        fecha=datetime.now().date())
                    preinscripcion.save(request)
            EXCLUDES.extend(mejores)
            """LISTADO DE DISCAPACITADOS EXCLUYENDO A LOS DE MEJORES PROMEDIOS"""
            discapacitados = lista_discapacitado_beca(periodoactual=periodoactual, periodoanterior=periodoanterior, excludes=EXCLUDES)
            inscripciones_discapacitados = Inscripcion.objects.filter(pk__in=discapacitados).exclude(persona_id__in=PreInscripcionBeca.objects.values("inscripcion__persona_id").filter(periodo=periodoactual))
            actualizar_discapacitados = PreInscripcionBeca.objects.filter(becatipo=becatipos.filter(pk=ID_DISCAPACIDAD).first(), periodo=periodoactual).exclude(inscripcion__in=inscripciones_discapacitados)
            if DEBUG:
                actualizar_discapacitados.delete()
            for inscripcion in inscripciones_discapacitados:
                if not PreInscripcionBeca.objects.filter(inscripcion=inscripcion, periodo=periodoactual).exists():
                    preinscripcion = PreInscripcionBeca(inscripcion=inscripcion,
                                                        promedio=inscripcion.promedio,
                                                        becatipo=becatipos.filter(pk=ID_DISCAPACIDAD).first(),
                                                        periodo=periodoactual,
                                                        fecha=datetime.now().date())
                    preinscripcion.save(request)
            EXCLUDES.extend(discapacitados)
            """LISTADO DE DEPORTISTAS EXCLUYENDO A DISCAPACITADOS Y LOS DE MEJORES PROMEDIOS"""
            deportistas = lista_deportista_beca(periodoactual=periodoactual, periodoanterior=periodoanterior, excludes=EXCLUDES)
            inscripciones_deportistas = Inscripcion.objects.filter(pk__in=deportistas).exclude(persona_id__in=PreInscripcionBeca.objects.values("inscripcion__persona_id").filter(periodo=periodoactual))
            actualizar_deportistas = PreInscripcionBeca.objects.filter(becatipo=becatipos.filter(pk=ID_DEPORTISTA).first(), periodo=periodoactual).exclude(inscripcion__in=inscripciones_deportistas)
            if DEBUG:
                actualizar_deportistas.delete()
            for inscripcion in inscripciones_deportistas:
                if not PreInscripcionBeca.objects.filter(inscripcion=inscripcion, periodo=periodoactual).exists():
                    preinscripcion = PreInscripcionBeca(inscripcion=inscripcion,
                                                        promedio=inscripcion.promedio,
                                                        becatipo=becatipos.filter(pk=ID_DEPORTISTA).first(),
                                                        periodo=periodoactual,
                                                        fecha=datetime.now().date())
                    preinscripcion.save(request)
            EXCLUDES.extend(deportistas)
            """LISTADO DE PERSONAS EN EL EXTRANJERO EXCLUYENDO A DISCAPACITADOS, DEPORTISTAS Y LOS DE MEJORES PROMEDIOS"""
            migrantes = lista_migrante_exterior_beca(periodoactual=periodoactual, periodoanterior=periodoanterior, excludes=EXCLUDES)
            inscripciones_migrantes = Inscripcion.objects.filter(pk__in=migrantes).exclude(persona_id__in=PreInscripcionBeca.objects.values("inscripcion__persona_id").filter(periodo=periodoactual))
            actualizar_migrantes = PreInscripcionBeca.objects.filter(becatipo=becatipos.filter(pk=ID_EXTERIOR_MIGRANTE).first(), periodo=periodoactual).exclude(inscripcion__in=inscripciones_migrantes)
            if DEBUG:
                actualizar_migrantes.delete()
            for inscripcion in inscripciones_migrantes:
                if not PreInscripcionBeca.objects.filter(inscripcion=inscripcion, periodo=periodoactual).exists():
                    preinscripcion = PreInscripcionBeca(inscripcion=inscripcion,
                                                        promedio=inscripcion.promedio,
                                                        becatipo=becatipos.filter(pk=ID_EXTERIOR_MIGRANTE).first(),
                                                        periodo=periodoactual,
                                                        fecha=datetime.now().date())
                    preinscripcion.save(request)
            EXCLUDES.extend(migrantes)
            """LISTADO DE PERSONAS ETNIA EXCLUYENDO EXTRANJERO, MIGRANTES, DISCAPACITADOS, DEPORTISTAS Y LOS DE MEJORES PROMEDIOS"""
            etnias = lista_etnia_beca(periodoactual=periodoactual, periodoanterior=periodoanterior, excludes=EXCLUDES)
            inscripciones_etnias = Inscripcion.objects.filter(pk__in=etnias).exclude(persona_id__in=PreInscripcionBeca.objects.values("inscripcion__persona_id").filter(periodo=periodoactual))
            actualizar_etnias = PreInscripcionBeca.objects.filter(becatipo=becatipos.filter(pk=ID_ETNIA).first(), periodo=periodoactual).exclude(inscripcion__in=inscripciones_etnias)
            if DEBUG:
                actualizar_etnias.delete()
            for inscripcion in inscripciones_etnias:
                if not PreInscripcionBeca.objects.filter(inscripcion=inscripcion, periodo=periodoactual).exists():
                    preinscripcion = PreInscripcionBeca(inscripcion=inscripcion,
                                                        promedio=inscripcion.promedio,
                                                        becatipo=becatipos.filter(pk=ID_ETNIA).first(),
                                                        periodo=periodoactual,
                                                        fecha=datetime.now().date())
                    preinscripcion.save(request)
            EXCLUDES.extend(etnias)

            # pruebas = PreInscripcionBeca.objects.filter(periodo=periodo).exclude(becatipo=becatipos.filter(pk=ID_GRUPO_VULNERABLE).first())
            # EXCLUDES.extend(list(pruebas.values_list('inscripcion_id', flat=True).distinct()))
            """LISTADO DE PERSONAS DE GRUPO VULNERABLE EXCLUYENDO ETNIA, EXTRANJERO, MIGRANTES, DISCAPACITADOS, DEPORTISTAS Y LOS DE MEJORES PROMEDIOS"""
            grupo = []
            for grupo_id in [4, 5]:
                grupo.extend(lista_gruposocioeconomico_beca(periodoactual=periodoactual, periodoanterior=periodoanterior, tipogrupo_id=grupo_id, excludes=EXCLUDES, limit=20))
            inscripciones_grupo = Inscripcion.objects.filter(pk__in=grupo).exclude(persona_id__in=PreInscripcionBeca.objects.values("inscripcion__persona_id").filter(periodo=periodoactual))
            actualizar_grupo = PreInscripcionBeca.objects.filter(becatipo=becatipos.filter(pk=ID_GRUPO_VULNERABLE).first(), periodo=periodoactual).exclude(inscripcion__in=inscripciones_grupo)
            if DEBUG:
                actualizar_grupo.delete()
            for inscripcion in inscripciones_grupo:
                if not PreInscripcionBeca.objects.filter(inscripcion=inscripcion, periodo=periodoactual).exists():
                    preinscripcion = PreInscripcionBeca(inscripcion=inscripcion,
                                                        promedio=inscripcion.promedio,
                                                        becatipo=becatipos.filter(pk=ID_GRUPO_VULNERABLE).first(),
                                                        periodo=periodoactual,
                                                        fecha=datetime.now().date())
                    preinscripcion.save(request)
            error = False

        except Exception as ex:
            transaction.set_rollback(True)
            print(ex)
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            error = True
            mensaje_ex = ex.__str__()

    if error:
        notificacion = Notificacion(titulo=f"Error en el proceso de creación de precandidatos de becas",
                                    cuerpo=f"Ocurrio un error al generar precandidatos de becas del periodo académico {periodoactual.nombre}. {mensaje_ex}",
                                    destinatario=request.session['persona'],
                                    url="/adm_becas",
                                    content_type=None,
                                    object_id=None,
                                    prioridad=2,
                                    app_label=request.session['tiposistema'],
                                    fecha_hora_visible=datetime.now() + timedelta(days=2)
                                    )
        notificacion.save(request)
    else:
        notificacion = Notificacion(titulo=f"Proceso finalizado de creación de precandidatos de becas",
                                    cuerpo=f"Se genero correctamente el precandidatos de becas del periodo académico {periodoactual.nombre}",
                                    destinatario=request.session['persona'],
                                    url="/adm_becas",
                                    content_type=None,
                                    object_id=None,
                                    prioridad=2,
                                    app_label=request.session['tiposistema'],
                                    fecha_hora_visible=datetime.now() + timedelta(days=2)
                                    )
        notificacion.save(request)


def generate_precandidatos_beca_desviacion_standar_todas_carrerasnew(request, data, periodoactual, periodoanterior):
    error = False
    mensaje_ex = None
    inicio = time.time()
    fin_preseleccion = None
    sendemail = variable_valor('ENVIAR_CORREO_BECAS') if variable_valor('ENVIAR_CORREO_BECAS') else False
    lista_envio_masivo = []
    with transaction.atomic():
        try:
            lista_envio_masivo = generar_precandidatos_becas(periodoanterior, periodoactual, usuario_ejecuta_id=request.user.id, enviarcorreo=sendemail)
            fin_preseleccion = time.time()
            print("Tiempo generar preseleccionados con solicitud: ", fin_preseleccion - inicio)
            error = False
        except Exception as ex:
            transaction.set_rollback(True)
            print(ex)
            err = 'Error on line {}'.format(sys.exc_info()[-1].tb_lineno)
            error = True
            mensaje_ex = f'{err} {ex.__str__()}'
    # Se comento envio de correo masivo
    for x in lista_envio_masivo:
        tituloemail = x['tituloemail']
        data = x['data']
        list_email = x['list_email']
        plantilla = x['plantilla']

        send_html_mail(tituloemail,
                       plantilla,
                       data,
                       list_email,
                       [],
                       cuenta=CUENTAS_CORREOS[0][1]
                       )
        time.sleep(3)
    #fin_envio_correo = time.time()
    #print("Tiempo envio de correo de preseleccionados: ",  fin_envio_correo - fin_preseleccion)
    #print("DURACIÓN TOTAL DE PROCESO DE PRESELECCION DE BECADOS: ", fin_envio_correo - inicio)

    if error:
        notificacion = Notificacion(titulo=f"Error en el proceso de creación de precandidatos de becas",
                                    cuerpo=f"Ocurrio un error al generar precandidatos de becas del periodo académico {periodoactual.nombre}. {mensaje_ex}",
                                    destinatario=request.session['persona'],
                                    url="/adm_becas",
                                    content_type=None,
                                    object_id=None,
                                    prioridad=2,
                                    app_label=request.session['tiposistema'],
                                    fecha_hora_visible=datetime.now() + timedelta(days=2)
                                    )
        notificacion.save(request)
    else:
        notificacion = Notificacion(titulo=f"Proceso finalizado de creación de precandidatos de becas",
                                    cuerpo=f"Se genero correctamente el precandidatos de becas del periodo académico {periodoactual.nombre}",
                                    destinatario=request.session['persona'],
                                    url="/adm_becas?action=listadopredatabecados",
                                    content_type=None,
                                    object_id=None,
                                    prioridad=2,
                                    app_label=request.session['tiposistema'],
                                    fecha_hora_visible=datetime.now() + timedelta(days=2)
                                    )
        notificacion.save(request)


def generar_requisitos_preinscripcionesbecas(request, data, periodoactual, periodoanterior):
    error = False
    mensaje_errores = None
    id = int(encrypt(request.POST['id']))
    idc = int(encrypt(request.POST['idc']))
    becatipo = BecaTipo.objects.get(pk=id)
    configuracion = becatipo.becatipoconfiguracion_set.get(pk=idc)
    observacion_error = []
    if  configuracion:
        preinscripciones = PreInscripcionBeca.objects.filter(periodo=periodoactual, becatipo_id=becatipo)

        for preinscripcion in preinscripciones:
            # print(preinscripcion)
            # requistos_removido = preinscripcion.preinscripcionbecarequisito_set.all().exclude(detallerequisitobeca__in=configuracion.requisitosbecas.all())
            # # Remover requisitos que  cambiaron en la configuracion
            # for reqb in requistos_removido:
            #     reqb.status = False
            #     reqb.save(request)
            #     log(u'%Elimino requisito beca: %s' % reqb, request, "del")
            # Agregar o Editar requisitos de prebecas

            with transaction.atomic(using='default'):
                try:
                    for i, detallerequisitobeca in enumerate(configuracion.requisitosbecas.filter(visible=True, status=True)):
                        preinscripcionbecarequisito = preinscripcion.preinscripcionbecarequisito_set.filter(detallerequisitobeca=detallerequisitobeca).first()
                        mensaje, logaction = 'Edito', 'edit'
                        if not preinscripcionbecarequisito:
                            preinscripcionbecarequisito = PreInscripcionBecaRequisito(
                                preinscripcion=preinscripcion,
                                detallerequisitobeca=detallerequisitobeca
                            )
                            mensaje, logaction = 'Adiciono', 'add'
                        preinscripcionbecarequisito.status = True
                        preinscripcionbecarequisito.run()
                        preinscripcionbecarequisito.save(request)
                        persona = request.user.persona_set.first()
                        if preinscripcionbecarequisito.detallerequisitobeca.funcionejecutar:
                            historial = HistorialPreInscripcionBecaRequisito(preinscripcionbecarequisito=preinscripcionbecarequisito,
                                                                             observacion=f'{mensaje} requisito beca',
                                                                             persona=persona,
                                                                             cumplerequisito=preinscripcionbecarequisito.cumplerequisito)
                            historial.save(request)

                        log(u'%s requisito beca: %s' % (mensaje, preinscripcionbecarequisito), request, logaction)
                        print(f'----------------->{i + 1} ', preinscripcionbecarequisito)
                except Exception as ex:
                    transaction.set_rollback(True, using='default')
                    msg = str(ex)
                    observacion_error.append(f'\n El registro con id ({preinscripcion.id}) que pertenece a {preinscripcion.inscripcion.persona}  ocurrio el siguiente error {msg}')
    else:
        observacion_error.append(f"El tipo de beca {becatipo.nombre} no tiene un configuración asignada")

    if len(observacion_error) > 0:
        mensaje_errores = ' '.join(observacion_error)
        becatipo = BecaTipo.objects.get(pk=id)
        notificacion = Notificacion(titulo=f"Error en el proceso de creación o actualización de requisitos becas",
                                    cuerpo=f"Ocurrio un error en la creación o actualización de requisitos becas del periodo académico {periodoactual.nombre} en {configuracion}. que son loa suiguientes errores: {mensaje_errores}",
                                    destinatario=request.session['persona'],
                                    url="/adm_becas",
                                    content_type=None,
                                    object_id=None,
                                    prioridad=2,
                                    app_label=request.session['tiposistema'],
                                    fecha_hora_visible=datetime.now() + timedelta(days=2)
                                    )
        notificacion.save(request)
    else:
        notificacion = Notificacion(titulo=f"Proceso finalizado de creación o actualización de requisitos becas",
                                    cuerpo=f"Se genero correctamente la creación o actualización de requisitos becas del periodo académico {periodoactual.nombre} en {configuracion}",
                                    destinatario=request.session['persona'],
                                    url="/adm_becas",
                                    content_type=None,
                                    object_id=None,
                                    prioridad=2,
                                    app_label=request.session['tiposistema'],
                                    fecha_hora_visible=datetime.now() + timedelta(days=2)
                                    )
        notificacion.save(request)


def importar_archivopreinscriptos_becas(request, data, periodoactual, periodoanterior, noti=None):
    error = False
    mensaje_ex = None
    requisito_archivo = data['requisito_archivo']
    configuracion = requisito_archivo.becatipoconfiguracion
    becatipo = requisito_archivo.becatipoconfiguracion.becatipo
    archivo = requisito_archivo.archivo.path
    requisitos = configuracion.requisitosbecas
    persona = data['persona']
    observacion_error = []
    ids = data['ids']
    try:
        wb = load_workbook(filename=archivo)
        sheet = wb[wb.sheetnames[0]]
        datae = []
        print('id_preinscripon', 'id_detallerequisito', 'cumplerequisito', 'observacion')
        for row in range(2, sheet.max_row + 1):
            rows = []
            id_preinscripcion = None
            col_aux = 2
            # procesamiento de datos
            preinscripcion = None
            for col in range(1, sheet.max_column):
                if col != 1:
                    if col == col_aux:
                        preinscripcionbecarequisitos = PreInscripcionBecaRequisito.objects.filter(preinscripcion=preinscripcion, detallerequisitobeca=ids[col])
                        if preinscripcionbecarequisitos:
                            preinscripcionbecarequisito = preinscripcionbecarequisitos.first()
                            cumplerequisito = sheet.cell(row=row, column=col).value
                            observacion = sheet.cell(row=row, column=col + 1).value
                            if cumplerequisito in ['SI', 'NO']:
                                if preinscripcionbecarequisitos.filter(detallerequisitobeca__visible=True, detallerequisitobeca__funcionejecutar__isnull=True):
                                    with transaction.atomic(using='default'):
                                        try:
                                            rows.append([id_preinscripcion, ids[col], cumplerequisito, observacion])
                                            preinscripcionbecarequisito.cumplerequisito = True if cumplerequisito == 'SI' else False
                                            preinscripcionbecarequisito.save(request)
                                            log(u'Cambio estado  de requisito beca: %s el estado  es %s' % (preinscripcionbecarequisito, preinscripcionbecarequisito.cumplerequisito), request, "edit")
                                            historial = HistorialPreInscripcionBecaRequisito(preinscripcionbecarequisito=preinscripcionbecarequisito,
                                                                                             observacion=observacion if observacion else 'SIN OBSERVACIÓN',
                                                                                             persona=persona,
                                                                                             cumplerequisito=preinscripcionbecarequisito.cumplerequisito)
                                            historial.save(request)
                                            print(id_preinscripcion, '            ', ids[col], '                ', cumplerequisito, '            ', observacion)
                                        except Exception as ex:
                                            transaction.set_rollback(True, using='default')
                                            msg = str(ex)
                                            observacion_error.append(f'En id_preinscripcion({id_preinscripcion})  ocurrio el siguiente error {msg}')
                                            err = 'Error on line {}'.format(sys.exc_info()[-1].tb_lineno)
                                else:
                                    print(id_preinscripcion, '            ', ids[col], '                ', cumplerequisito, '            ', observacion, '------>Configurado por el sistema No se guarda')
                            else:
                                observacion_error.append(f'En el id_preinscripcion({id_preinscripcion})  digito mal la opción de cumplimiento. Escribió  {cumplerequisito} debe escribir (Si,NO)')
                        else:
                            message_error = f'No  existe  id_preinscripcion({id_preinscripcion}) en el tipo de beca  {becatipo} no existe el requisito {ids[col]}'
                            observacion_error.append(message_error)
                        col_aux += 2
                else:
                    id_preinscripcion = int(sheet.cell(row=row, column=col).value)
                    preinscripcion = PreInscripcionBeca.objects.filter(becatipo=becatipo, id=id_preinscripcion).first()
                    if not preinscripcion:
                        message_error = f'No  existe  id_preinscripcion({id_preinscripcion}) en el tipo de beca  {becatipo}'
                        observacion_error.append(message_error)
            datae.append(rows)
        error = True if observacion_error else False
    except Exception as ex:
            # transaction.set_rollback(True)
            print(ex)
            err = 'Error on line {}'.format(sys.exc_info()[-1].tb_lineno)
            error = True
            mensaje_ex = f'{err} {ex.__str__()}'
            observacion_error.append(mensaje_ex)

    notificacion = Notificacion(
        destinatario=request.session['persona'],
        url="/adm_becas",
        content_type=None,
        object_id=None,
        prioridad=2,
        app_label=request.session['tiposistema'],
        fecha_hora_visible=datetime.now() + timedelta(days=2)
        )

    if noti:
        notificacion = noti
        notificacion.prioridad = 2

    if error:
        requisito_archivo.estado = 3
        requisito_archivo.observacion_error = str(observacion_error)
        notificacion.titulo = f"Error en el proceso de creación o actualización de requisitos becas"
        notificacion.cuerpo = f"Ocurrio un error en la creación o actualización de requisitos becas del periodo académico {periodoactual.nombre} en {configuracion}. errores ocurridos en la ejecuciòn {observacion_error}"
    else:
        requisito_archivo.estado = 2
        notificacion.titulo = f"Proceso finalizado de creación o actualización de requisitos becas"
        notificacion.cuerpo = f"Se genero correctamente la creación o actualización de requisitos becas del periodo académico {periodoactual.nombre} en {configuracion}"

    requisito_archivo.save(request)
    notificacion.save(usuario_id=1)


def importar_archivoexcelbecarios(request, data, periodo, periodovalida, noti=None, enviarcorreo=False):
    error = False
    mensaje_ex = None
    lista_envio_masivo = []
    errores = []
    personaejecuta = data['persona']
    archivobecarios = data['archivobecarios']
    becatipo = data['becatipo']
    try:
        wb = load_workbook(filename=archivobecarios.archivo.path, read_only=True)
        sheet = wb[wb.sheetnames[0]]
        ids = {}
        for row in range(2, sheet.max_row + 1):
            prei_id = sheet.cell(row=row, column=1).value
            preinscripcionbeca = PreInscripcionBeca.objects.filter(pk=prei_id).first()
            if preinscripcionbeca is not None:#Validación de preinscripcion si exista.
                persona = preinscripcionbeca.inscripcion.persona
                if preinscripcionbeca.becatipo.id == becatipo.id:#Validación de tipo de beca de que acepta.
                    if preinscripcionbeca.cumple_requisitos():#Validación de  que cumpla todos los requisitos.
                        requisitos = BecaRequisitos.objects.filter(Q(becatipo__isnull=False),
                                                                    becatipo=preinscripcionbeca.becatipo,
                                                                    periodo=periodo, status=True,
                                                                    vigente=True).order_by('numero')
                        if requisitos.exists():
                            becado = BecaSolicitud.objects.filter(inscripcion=preinscripcionbeca.inscripcion,
                                                                  becatipo=preinscripcionbeca.becatipo,
                                                                  periodo=periodo,
                                                                  periodocalifica=periodovalida,
                                                                  status=True).first()
                            if becado is None:
                                with transaction.atomic(using='default'):
                                    try:
                                        becado = BecaSolicitud(
                                            inscripcion=preinscripcionbeca.inscripcion,
                                            becatipo=preinscripcionbeca.becatipo,
                                            periodo=periodo,
                                            periodocalifica=periodovalida,
                                            estado=1,
                                            tiposolicitud=preinscripcionbeca.tipo_renovacion_nueva(periodovalida),
                                            observacion=f'SEMESTRE REGULAR {periodo.nombre}')
                                        becado.save(request)
                                        log(u'Agrego solicitud de beca: %s' % becado, request, "add")
                                        #asignar requisitos validados por PreincripcionBeca
                                        for requisito in requisitos:
                                            becadetallesolicitud = BecaDetalleSolicitud.objects.filter(status=True, solicitud=becado,
                                                                                                       requisito=requisito, cumple=True,
                                                                                                       estado=2).first()
                                            if becadetallesolicitud is None:
                                                becadetallesolicitud = BecaDetalleSolicitud(solicitud=becado,
                                                                                            requisito=requisito,
                                                                                            cumple=True,
                                                                                            archivo=None,
                                                                                            estado=2,
                                                                                            observacion='APROBACIÓN AUTOMÁTICA',
                                                                                            personaaprueba=None,
                                                                                            fechaaprueba=None)
                                                becadetallesolicitud.save(request)
                                                log(u'Agrego detalle solicitud de baca: %s' % becadetallesolicitud, request, "add")


                                        recorrido = BecaSolicitudRecorrido.objects.filter(solicitud=becado, estado=1).first()
                                        if recorrido is None:
                                            recorrido = BecaSolicitudRecorrido(solicitud=becado,
                                                                               observacion="SOLICITUD AUTOMÁTICA",
                                                                               estado=1)
                                            recorrido.save(request)#usuario_id=1
                                            log(u'Agrego recorrido de beca: %s' % recorrido, request, "add")
                                            # notificacion = Notificacion(titulo=f"Registro de Solicitud de Beca - {periodo.nombre}",
                                            #                             cuerpo=f"Registró de solicitud de aplicación a la beca por {becado.becatipo.nombre.upper()} para el periodo académico {periodo.nombre} el {datetime.now().strftime('%d-%m-%Y')} a las {datetime.now().strftime('%H:%M %p')}",
                                            #                             destinatario=persona,
                                            #                             url="/alu_becas",
                                            #                             content_type=None,
                                            #                             object_id=None,
                                            #                             prioridad=2,
                                            #                             perfil=preinscripcionbeca.inscripcion.perfil_usuario(),
                                            #                             app_label='SIE',#request.session['tiposistema'],
                                            #                             fecha_hora_visible=datetime.now() + timedelta(days=2)
                                            #                             )
                                            # notificacion.save(usuario_id=1)
                                            # if enviarcorreo:
                                            #
                                            #     tituloemail = f"Registro de Solicitud de Beca - {periodo.nombre}"
                                            #     data = {'sistema': u'SGA - UNEMI',
                                            #             'tipo': 'SOL',
                                            #             'tipobeca': becado.becatipo.nombre.upper(),
                                            #             'fecha': datetime.now().date(),
                                            #             'hora': datetime.now().time(),
                                            #             'periodo': becado.periodo.nombre,
                                            #             'estudiante': persona,
                                            #             'autoridad2': '',
                                            #             't': miinstitucion()
                                            #             }
                                            #     list_email = persona.lista_emails_envio()
                                            #     if DEBUG:
                                            #         list_email = ['atorrese@unemi.edu.ec']
                                            #     lista_envio_masivo.append({'tituloemail': tituloemail,
                                            #                                'data': data,
                                            #                                'list_email': list_email,
                                            #                                'plantilla': "emails/solicitudbecaestudiante.html"})

                                        recorrido = BecaSolicitudRecorrido.objects.filter(solicitud=becado, estado=4).first()
                                        # REGISTRO EN ESTADO DE REVISION
                                        if recorrido is None:
                                            becado.estado = 4
                                            becado.save(request)
                                            recorrido = BecaSolicitudRecorrido(solicitud=becado,
                                                                               observacion="EN REVISION",
                                                                               estado=4)
                                            recorrido.save(usuario_id=1)
                                            log(u'Agrego recorrido de beca: %s' % recorrido, request, "add")

                                        recorrido = BecaSolicitudRecorrido.objects.filter(solicitud=becado, estado=2).first()
                                        if recorrido is None:
                                            becado.estado = 2
                                            becado.becaaceptada = 1
                                            becado.save(request)
                                            recorrido = BecaSolicitudRecorrido(solicitud=becado,
                                                                               observacion="PENDIENTE DE ACEPTACIÓN O RECHAZO",
                                                                               estado=2)
                                            recorrido.save(request)
                                            log(u'Agrego recorrido de beca: %s' % recorrido, request, "add")
                                            notificacion = Notificacion(titulo=f"Solicitud de Beca en Revisión - {periodo.nombre}",
                                                                        cuerpo=f"Solicitud de aplicación a la beca por {becado.becatipo.nombre.upper()} para el periodo académico {becado.periodo.nombre} ha sido {'APROBADA' if becado.estado == 2 else 'RECHAZADA'}",
                                                                        destinatario=persona,
                                                                        url="/alu_becas",
                                                                        content_type=None,
                                                                        object_id=None,
                                                                        prioridad=2,
                                                                        perfil=preinscripcionbeca.inscripcion.perfil_usuario(),
                                                                        app_label='SIE',#request.session['tiposistema'],
                                                                        fecha_hora_visible=datetime.now() + timedelta(days=2)
                                                                        )
                                            notificacion.save(usuario_id=1)
                                            # Envio de e-mail de notificación al solicitante
                                            if enviarcorreo:
                                                tituloemail = "Beca estudiantil"
                                                data = {'sistema': u'SGA - UNEMI',
                                                        'fase': 'AR',
                                                        'tipobeca': becado.becatipo.nombre.upper(),
                                                        'fecha': datetime.now().date(),
                                                        'hora': datetime.now().time(),
                                                        'saludo': 'Estimada' if becado.inscripcion.persona.sexo_id == 1 else 'Estimado',
                                                        'estado': 'APROBADA' if becado.estado == 2 else "RECHAZADA",
                                                        'estudiante': persona.nombre_minus(),
                                                        'autoridad2': '',
                                                        'observaciones': '',
                                                        'periodo': periodo.nombre,
                                                        't': miinstitucion()
                                                        }
                                                list_email = persona.lista_emails_envio()
                                                if DEBUG:
                                                    list_email = ['atorrese@unemi.edu.ec']
                                                lista_envio_masivo.append({'tituloemail': tituloemail,
                                                                           'data': data,
                                                                           'list_email': list_email,
                                                                           'plantilla': "emails/notificarestadosolicitudbeca.html"})

                                        print(f"------ Solicitud Becado registrado exitosamente: ", persona)
                                    except Exception as ex:
                                        transaction.set_rollback(True, using='default')
                                        msg = str(ex)
                                        errores.append(f'En ID_PREINSCRIPCION({prei_id})  en la fila {row}  ocurrio el siguiente error {msg}')
                                        err = 'Error on line {}'.format(sys.exc_info()[-1].tb_lineno)
                            else:
                                errores.append(f'El el ID_PREINSCRIPCION({prei_id}) de la persona:({persona}) en la fila {row} ya tiene una solicitud beca registrada')
                        else:
                            errores.append(f'El el ID_PREINSCRIPCION({prei_id}) de la persona:({persona}) en la fila {row} no existe configuración de requisitos para crear solicitud')
                    else:
                        errores.append(f'El el ID_PREINSCRIPCION({prei_id}) de la persona:({persona}) en la fila {row} tiene uno o varios requisitos que  no cumple')
                else:
                    errores.append(f'El el ID_PREINSCRIPCION({prei_id}) de la persona:({persona}) en la fila {row} tiene el tipo de beca({preinscripcionbeca.becatipo}) solo permite el tipo {becatipo})')
            else:
                errores.append(f'No existe el ID_PREINSCRIPCION({prei_id}) en la fila {row}')


        print("Proceso terminado...")

    except Exception as ex:
        print(ex)
        mensaje_ex = f'Error on line {sys.exc_info()[-1].tb_lineno} {ex.__str__()}'
        print(mensaje_ex)
        errores.append(mensaje_ex)
        lista_envio_masivo = []

    for x in lista_envio_masivo:
        tituloemail = x['tituloemail']
        data = x['data']
        list_email = x['list_email']
        plantilla = x['plantilla']

        send_html_mail(tituloemail,
                       plantilla,
                       data,
                       list_email,
                       [],
                       cuenta=CUENTAS_CORREOS[0][1]
                       )
        time.sleep(5)

    notificacion = Notificacion(
        destinatario=request.session['persona'],
        url="/adm_becas",
        content_type=None,
        object_id=None,
        prioridad=2,
        app_label='SIE',#request.session['tiposistema'],
        fecha_hora_visible=datetime.now() + timedelta(days=2)
        )

    if noti:
        notificacion = noti
        notificacion.prioridad = 2

    if errores:
        archivobecarios.estado = 3
        archivobecarios.observacion_error = str(errores)
        #archivobecarios.archivo
        notificacion.titulo = f"Proceso de Importación archivo becarios contiene errores",
        notificacion.cuerpo = f'El archivo contiene los siguientes errores => {archivobecarios.observacion_error}'
    else:
        archivobecarios.estado = 2
        notificacion.titulo = f"Proceso de Importación archivo becarios finalizado exitosamente"

    archivobecarios.save(request)
    notificacion.save(usuario_id=1)


def corregir_actascompromisos_becas(request, data, periodo,  noti=None):
    aData = {}
    errores = []
    url_path = 'http://127.0.0.1:8000'
    if not DEBUG:
        url_path = 'https://sga.unemi.edu.ec'
    becas_solicitudes = BecaSolicitud.objects.filter(periodo_id=periodo.id,
                                                     becaaceptada=2,
                                                     archivoactacompromiso='',
                                                     status=True)
    if DEBUG:
        becas_solicitudes = becas_solicitudes[:10]

    for key, becasolicitud in enumerate(becas_solicitudes):
        with transaction.atomic():
            try:
                aData = {}
                persona_beca = becasolicitud.inscripcion.persona
                becaasignada = becasolicitud.becaasignacion_set.filter(status=True).first()
                if becaasignada:
                    output_folder = ''
                    aData['becasolicitud'] = becasolicitud
                    aData['documentopersonal'] = documentopersonal = persona_beca.personadocumentopersonal_set.filter(status=True).first()
                    aData['cuentabancaria'] = cuentabancaria = cuentabancaria = persona_beca.cuentabancaria_becas()
                    aData['perfil'] = perfil = persona_beca.mi_perfil()
                    aData['deportista'] = deportista = persona_beca.deportistapersona_set.filter(status=True).first()
                    aData['isPersonaExterior'] = isPersonaExterior = persona_beca.ecuatoriano_vive_exterior()
                    aData['configuracionbecatipoperiodo'] = becatipoconfiguracion = becasolicitud.obtener_configuracionbecatipoperiodo()
                    aData['matricula'] = matricula = becasolicitud.obtener_matricula()
                    eInscripcion = becasolicitud.inscripcion
                    ePeriodo = becasolicitud.periodo
                    eUsuario = persona_beca.usuario
                    username = elimina_tildes(eUsuario.username)
                    filename = f'acta_compromiso_{eInscripcion.id}_{ePeriodo.id}_{becasolicitud.id}'
                    filenametemp = os.path.join(os.path.join(SITE_STORAGE, 'media', 'becas', 'temp', 'actas_compromisos', username, filename + '.pdf'))
                    filenameqrtemp = os.path.join(os.path.join(SITE_STORAGE, 'media', 'becas', 'temp', 'actas_compromisos', username, 'qrcode', filename + '.png'))
                    url_actafirmada = None
                    folder = os.path.join(os.path.join(SITE_STORAGE, 'media', 'becas', 'actas_compromisos', username, ''))
                    folder2 = os.path.join(os.path.join(SITE_STORAGE, 'media', 'becas', 'actas_compromisos', username, 'qrcode', ''))
                    aData['aceptobeca'] = True
                    aData['url_qr'] = rutaimg = folder2 + filename + '.png'
                    aData['rutapdf'] = rutapdf = folder + filename + '.pdf'
                    aData['url_pdf'] = url_pdf = f'{url_path}/media/becas/actas_compromisos/{username}/{filename}.pdf'
                    if os.path.isfile(rutapdf):
                        os.remove(rutapdf)
                    elif os.path.isfile(rutaimg):
                        os.remove(rutaimg)
                    os.makedirs(folder, exist_ok=True)
                    os.makedirs(folder2, exist_ok=True)
                    firma = f'ACEPTADO POR: {eInscripcion.persona.__str__()}\nUSUARIO:{username}\nFECHA: {becaasignada.fecha_creacion.__str__()}\nACEPTO EN: sga.unemi.edu.ec\nDOCUMENTO:{url_pdf}'
                    url = pyqrcode.create(firma)
                    imageqr = url.png(rutaimg, 16, '#000000')
                    aData['version'] = datetime.now().strftime('%Y%m%d_%H%M%S')
                    aData['image_qrcode'] = f'{url_path}/media/becas/actas_compromisos/{username}/qrcode/{filename}.png'
                    aData['fechaactual'] = becaasignada.fecha_creacion  # datetime.now()
                    url_acta = f'becas/actas_compromisos/{username}/{filename}.pdf'
                    rutapdf = folder + filename + '.pdf'

                    valida, pdf, result = conviert_html_to_pdfsaveqr_generico(request, 'alu_becas/actav2.html', {
                        'pagesize': 'A4',
                        'data': aData,

                    }, folder, filename + '.pdf')
                    if not valida:
                        raise NameError(f'Error al generar el pdf de acta de compromiso {becasolicitud}')

                    becasolicitud.archivoactacompromiso = url_acta
                    becasolicitud.save(request)
                    time.sleep(3)
            except Exception as ex1:
                errores.append(ex1.__str__())
                transaction.set_rollback(True)
                print(ex1)
    notificacion = Notificacion(
        destinatario=request.session['persona'],
        url="/adm_becas",
        content_type=None,
        object_id=None,
        prioridad=2,
        app_label='SGA',  # request.session['tiposistema'],
        fecha_hora_visible=datetime.now() + timedelta(days=2)
    )

    if noti:
        notificacion = noti
        notificacion.prioridad = 2

    if errores:
        # archivobecarios.archivo
        notificacion.titulo = f"Proceso de corrección acta de compromiso becarios contiene errores",
        cuerpo = "\n".join(errores)
        notificacion.cuerpo = f'El archivo contiene los siguientes errores => {cuerpo}'
    else:
        notificacion.titulo = f"Proceso de corrección acta de compromiso becarios finalizado exitosamente"
    notificacion.save(usuario_id=1)


def generate_reporte_pendientes_pago_becas_financiero(request, data, periodo, noti=None):
    nombre_archivo = 'reporte_pagos_pendientes' + random.randint(1, 10000).__str__() + '.csv'
    folder = os.path.join(MEDIA_ROOT, 'becas', 'pagos_pendientes_financiero', '')
    os.makedirs(folder, exist_ok=True)
    directory = f'{folder}{nombre_archivo}'
    usernotify = request.user
    personadestino = Persona.objects.get(usuario_id=request.user.pk)
    with transaction.atomic():
        try:
            personadestino = request.session['persona']
            workbook = xlsxwriter.Workbook(directory)
            ws = workbook.add_worksheet('DETALLE')
            formatocabeceracolumna = workbook.add_format({
                'bold': 1,
                'border': 1,
                'align': 'center',
                'valign': 'vcenter',
                'bg_color': 'silver',
                'text_wrap': 1,
                'font_size': 10})

            formatocelda = workbook.add_format({
                'border': 1
            })

            formatotitulo = workbook.add_format(
                {'align': 'center', 'valign': 'vcenter', 'bold': 1, 'font_size': 16})
            columns = [
                (u"IDENTIFICACIÓN", 15),
                (u"ESTUDIANTE", 50),
                (u"VALOR", 10),
                (u"TIPO BECA", 50),
            ]
            row_num = 0
            for col_num in range(len(columns)):
                ws.write(row_num, col_num, columns[col_num][0], formatocabeceracolumna)
                ws.set_column(col_num, col_num, columns[col_num][1])
            pagos = SolicitudPagoBecaDetalle.objects.filter(solicitudpago__periodo=periodo,
                                                            status=True,
                                                            pagado=False,
                                                            generadofinanciero=False)
            row_num = 1
            for pago in pagos:
                ws.write(row_num, 0, pago.asignacion.solicitud.inscripcion.persona.identificacion(), formatocelda)
                ws.write(row_num, 1, remover_caracteres_tildes_unicode(pago.asignacion.solicitud.inscripcion.persona.__str__()), formatocelda)
                ws.write(row_num, 2, pago.monto, formatocelda)
                ws.write(row_num, 3, remover_caracteres_tildes_unicode(pago.asignacion.solicitud.becatipo.__str__()), formatocelda)
                pago.generadofinanciero = True
                pago.save(request)
                row_num += 1
            workbook.close()
            response = HttpResponse(directory, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=%s' % nombre_archivo
            if noti is None:
                noti = Notificacion(cuerpo='Reporte Listo',
                                    titulo='Reporte de pendiente de pagos beca financiero',
                                    destinatario=personadestino,
                                    url="{}becas/pagos_pendientes_financiero/{}".format(MEDIA_URL, nombre_archivo),
                                    prioridad=1,
                                    app_label='SGA',
                                    fecha_hora_visible=datetime.now() + timedelta(days=1), tipo=2, en_proceso=False)
                noti.save(request)
            else:
                noti.en_proceso = False
                noti.cuerpo = 'Reporte Listo'
                noti.url = "{}becas/pagos_pendientes_financiero/{}".format(MEDIA_URL, nombre_archivo)
                noti.save()
            send_user_notification(user=usernotify, payload={
                "head": "Reporte terminado",
                "body": 'Reporte de Pendientes de Pagos Beca Financiero',
                "action": "notificacion",
                "timestamp": time.mktime(datetime.now().timetuple()),
                "url": "{}becas/pagos_pendientes_financiero/{}".format(MEDIA_URL, nombre_archivo),
                "btn_notificaciones": traerNotificaciones(request, data, personadestino),
                "mensaje": "Su reporte ha sido terminado"
            }, ttl=500)
        except Exception as ex:
            transaction.set_rollback(True)
            print(ex)
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            textoerror = 'Reporte Fallido - {} Linea:{}'.format(ex, sys.exc_info()[-1].tb_lineno)
            if noti is None:
                noti = Notificacion(cuerpo='Reporte Fallido', titulo='Reporte de Pendientes de Pagos Beca Financiero falló en la ejecución',
                                    destinatario=personadestino, prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1), tipo=2, en_proceso=False, error=True)
                noti.save(request)
            else:
                noti.en_proceso = False
                noti.error = True
                noti.titulo = 'Reporte de Pendientes de Pagos Beca Financiero falló en la ejecución'
                noti.cuerpo = textoerror
                # noti.url = "{}reportes/prematricula/{}".format(MEDIA_URL, nombre_archivo)
                noti.save()

            send_user_notification(user=usernotify, payload={
                "head": "Reporte Fallido",
                "body": 'Reporte de Pendientes de Pagos Beca Financiero a fallado',
                "action": "notificacion",
                "timestamp": time.mktime(datetime.now().timetuple()),
                "btn_notificaciones": traerNotificaciones(request, data, personadestino),
                "mensaje": textoerror,
                "error": True
            }, ttl=500)
    # nombre_archivo = 'reporte_total_inscritos_facultad_capacitacion_docente' + random.randint(1, 10000).__str__() + '.xls'
    # directory = os.path.join(MEDIA_ROOT, 'reportes', 'capdocente', nombre_archivo)
    # try:

def enviarcorreo_precandidatos_beca_todas_carreras(request, data, periodoactual, periodoanterior, noti=None):
    try:
        error = False
        mensaje_ex = None
        inicio = time.time()
        fin_preseleccion = None
        tituloemail = "Beca estudiantil"
        lista_envio_masivo = []
        becas = BecaSolicitud.objects.filter(periodo=periodoactual, periodocalifica=periodoanterior, becaaceptada=1, status=True, notificado=False)
        for becado in becas:
            noti_estudiante = Notificacion(titulo=f"Solicitud de Beca en Revisión - {becado.periodo.nombre}",
                                        cuerpo=f"Solicitud de aplicación a la beca por {becado.becatipo.nombre.upper()} para el periodo académico {becado.periodo.nombre} ha sido {'APROBADA' if becado.estado == 2 else 'RECHAZADA'}",
                                        destinatario=becado.inscripcion.persona,
                                        url="/alu_becas",
                                        content_type=None,
                                        object_id=None,
                                        prioridad=2,
                                        perfil=becado.inscripcion.perfil_usuario(),
                                        app_label='SIE',  # request.session['tiposistema'],
                                        fecha_hora_visible=datetime.now() + timedelta(days=2)
                                        )
            noti_estudiante.save(usuario_id=1)
            context = {'sistema': u'SGA - UNEMI',
                    'fase': 'AR',
                    'tipobeca': becado.becatipo.nombre.upper(),
                    'fecha': datetime.now().date(),
                    'hora': datetime.now().time(),
                    'saludo': 'Estimada' if becado.inscripcion.persona.sexo_id == 1 else 'Estimado',
                    'estado': 'APROBADA' if becado.estado == 2 else "RECHAZADA",
                    'estudiante': becado.inscripcion.persona.nombre_minus(),
                    'autoridad2': '',
                    'observaciones': '',
                    'periodo': becado.periodo.nombre,
                    't': miinstitucion()
                    }
            list_email = becado.inscripcion.persona.lista_emails_envio()
            send_html_mail(tituloemail,
                           'emails/notificarestadosolicitudbeca.html',
                           context,
                           list_email,
                           [],
                           cuenta=CUENTAS_CORREOS[0][1]
                           )
            becado.notificado=True
            becado.save(update_fields=['notificado'])
            time.sleep(3)
        if not noti:
            notificacion = Notificacion(titulo=f"Envio de correos a precandidatos de becas exitoso",
                                        cuerpo=f"Se envio {len(becas)} correos de notificado a precandidatos de becas exitosamente",
                                        destinatario=request.session['persona'],
                                        url="/adm_becas?action=listadopredatabecados",
                                        content_type=None,
                                        object_id=None,
                                        prioridad=2,
                                        app_label=request.session['tiposistema'],
                                        fecha_hora_visible=datetime.now() + timedelta(days=2)
                                        )
            notificacion.save(request)
        else:
            noti.titulo=f"Envio de correos a precandidatos de becas exitoso"
            noti.cuerpo=f"Se envio {len(becas)} correos de notificado a precandidatos de becas exitosamente"
            noti.destinatario=request.session['persona']
            noti.url="/adm_becas?action=listadopredatabecados"
            noti.content_type=None
            noti.object_id=None
            noti.prioridad=2
            noti.app_label=request.session['tiposistema']
            noti.fecha_hora_visible=datetime.now() + timedelta(days=2)
            noti.save(request)
    except Exception as ex:
        transaction.set_rollback(True)
        print(ex)
        err = 'Error on line {}'.format(sys.exc_info()[-1].tb_lineno)
        error = True
        mensaje_ex = f'{err} {ex.__str__()}'
        if noti:
            noti.status=False
            notif.save()
        notificacion = Notificacion(titulo=f"Error en el envio de correo de becas",
                                    cuerpo=f"Ocurrio un error al notificar a los precandidatos del periodo: {periodoactual.nombre}. {mensaje_ex}",
                                    destinatario=request.session['persona'],
                                    url="/adm_becas?action=listadopredatabecados",
                                    content_type=None,
                                    object_id=None,
                                    prioridad=2,
                                    app_label=request.session['tiposistema'],
                                    fecha_hora_visible=datetime.now() + timedelta(days=2)
                                    )
        notificacion.save(request)

def anular_beca(request, beca, form):
    try:
        becasolicitud = beca.solicitud
        obs = form.cleaned_data['observacion'].strip().upper()
        ePersona = becasolicitud.inscripcion.persona
        becatipoconfiguracion = becasolicitud.obtener_configuracionbecatipoperiodo()
        beca.estadobeca = 3
        beca.status = False
        beca.save(request)

        becasolicitud.observacion = '**'+becasolicitud.observacion +'**'
        becasolicitud.becaaceptada = 3
        becasolicitud.save(request)

        detalle = BecaSolicitudRecorrido(solicitud=becasolicitud,
                                         estado=7,
                                         observacion=form.cleaned_data['observacion'].strip().upper(),
                                         fecha=datetime.now().date()
                                         )
        detalle.save(request)
        if (eBecaPeriodo := becasolicitud.periodo.becaperiodo_set.filter(status=True).first()) is not None:
            # Envio de e-mail de notificación al solicitante
            tituloemail = "Beca Anulada"
            send_html_mail(tituloemail,
                           "emails/notificaranulacionbeca.html",
                           {'sistema': u'SGA - UNEMI',
                            'fecha': datetime.now().date(),
                            'hora': datetime.now().time(),
                            'saludo': 'Estimada' if becasolicitud.inscripcion.persona.sexo_id == 1 else 'Estimado',
                            'estudiante': ePersona.nombre_completo_inverso(),
                            'periodo': becasolicitud.periodo.__str__(),
                            'autoridad2': '',
                            'observaciones': obs,
                            't': miinstitucion()
                            },
                           ePersona.lista_emails_envio(),
                           [],
                           cuenta=variable_valor('CUENTAS_CORREOS')[0]
                           )
        log(f'Anulo la beca estudiantil: {ePersona} | Motivo: {obs}', request, "edit")
        llenar_solicitudes_precandidatos(request, becasolicitud, becatipoconfiguracion)
    except Exception as ex:
        raise NameError(str(ex))

def llenar_solicitudes_precandidatos(request, becasolicitud, becatipoconfiguracion):
    # AL RECHAZAR EL ESTUDIANTE SU BECA SE SELECCIONARA DE LA PREDATA LOS SIGUIENTES BECARIOS QUE NO FUERON SELECCIONADOS
    cantidad_estudiantes_becados = BecaSolicitud.objects.values('id').filter(status=True, periodo_id=becasolicitud.periodo_id).exclude(becaaceptada=3).count()
    cantidad_limite_becados = becatipoconfiguracion.becaperiodo.limitebecados
    preinscripcion_rechazo = PreInscripcionBeca.objects.filter(inscripcion=becasolicitud.inscripcion, status=True, periodo=becasolicitud.periodo).first()
    becados_rechazados = BecaSolicitud.objects.values('inscripcion_id').filter(periodo_id=becasolicitud.periodo_id, becaaceptada=3)

    while cantidad_estudiantes_becados < cantidad_limite_becados:
        preinscripcionbeca = PreInscripcionBeca.objects.filter(seleccionado=False, status=True, periodo_id=becasolicitud.periodo_id).exclude(
            inscripcion_id__in=becados_rechazados).order_by('orden').first()
        if (preinscripcionbeca_p := PreInscripcionBeca.objects.filter(seleccionado=False, status=True, prioridad=True, periodo_id=becasolicitud.periodo_id).exclude(
                inscripcion_id__in=becados_rechazados).order_by('orden').first()) is not None:
            preinscripcionbeca = preinscripcionbeca_p
        if preinscripcionbeca and not BecaSolicitud.objects.filter(status=True, inscripcion=preinscripcionbeca.inscripcion, periodo=becasolicitud.periodo).exists():
            becado = BecaSolicitud(inscripcion=preinscripcionbeca.inscripcion,
                                   becatipo=preinscripcionbeca.becatipo,
                                   periodo=becasolicitud.periodo,
                                   periodocalifica=becasolicitud.periodocalifica,
                                   estado=1,
                                   tiposolicitud=preinscripcionbeca.tipo_renovacion_nueva(becasolicitud.periodocalifica),
                                   observacion=f'SEMESTRE REGULAR {preinscripcionbeca.periodo.nombre}')
            becado.save(usuario_id=1)
            recorrido = BecaSolicitudRecorrido.objects.filter(solicitud=becado, estado=1).first()
            preinscripcionbeca.seleccionado = True
            preinscripcionbeca.save(usuario_id=1)
            if recorrido is None:
                recorrido = BecaSolicitudRecorrido(solicitud=becado,
                                                   observacion="SOLICITUD AUTOMÁTICA",
                                                   estado=1)
                recorrido.save(usuario_id=1)
                recorrido = BecaSolicitudRecorrido.objects.filter(solicitud=becado, estado=4).first()
                # REGISTRO EN ESTADO DE REVISION
                if recorrido is None:
                    becado.estado = 4
                    becado.save(usuario_id=1)
                    recorrido = BecaSolicitudRecorrido(solicitud=becado,
                                                       observacion="EN REVISION",
                                                       estado=4)
                    recorrido.save(usuario_id=1)
                    # log(u'Agrego recorrido de beca: %s' % recorrido, request, "add")

                recorrido = BecaSolicitudRecorrido.objects.filter(solicitud=becado, estado=2).first()
                if recorrido is None:
                    becado.estado = 2
                    becado.becaaceptada = 1
                    becado.save(usuario_id=1)
                    recorrido = BecaSolicitudRecorrido(solicitud=becado,
                                                       observacion="PENDIENTE DE ACEPTACIÓN O RECHAZO",
                                                       estado=2)
                    recorrido.save(usuario_id=1)
                    # log(u'Agrego recorrido de beca: %s' % recorrido, request, "add")
                    notificacion = Notificacion(titulo=f"Solicitud de Beca en Revisión - {preinscripcionbeca.periodo.nombre}",
                                                cuerpo=f"Solicitud de aplicación a la beca por {becado.becatipo.nombre.upper()} para el periodo académico {becado.periodo.nombre} ha sido {'APROBADA' if becado.estado == 2 else 'RECHAZADA'}",
                                                destinatario=preinscripcionbeca.inscripcion.persona,
                                                url="/alu_becas",
                                                content_type=None,
                                                object_id=None,
                                                prioridad=2,
                                                perfil=preinscripcionbeca.inscripcion.perfil_usuario(),
                                                app_label='SIE',  # request.session['tiposistema'],
                                                fecha_hora_visible=datetime.now() + timedelta(days=2)
                                                )
                    notificacion.save(usuario_id=1)
                    preinscripcion_rechazo.seleccionado = False
                    preinscripcion_rechazo.save(request)
                    tituloemail = "Beca estudiantil"
                    data = {'sistema': u'SGA - UNEMI',
                            'fase': 'AR',
                            'tipobeca': preinscripcionbeca.becatipo.nombre.upper(),
                            'fecha': datetime.now().date(),
                            'hora': datetime.now().time(),
                            'saludo': 'Estimada' if becado.inscripcion.persona.sexo_id == 1 else 'Estimado',
                            'estado': 'APROBADA' if becado.estado == 2 else "RECHAZADA",
                            'estudiante': becado.inscripcion.persona.nombre_minus(),
                            'autoridad2': '',
                            'observaciones': '',
                            'periodo': becado.periodo.nombre,
                            't': miinstitucion()
                            }
                    list_email = becado.inscripcion.persona.lista_emails_envio()
                    plantilla = "emails/notificarestadosolicitudbeca.html"
                    send_html_mail(tituloemail,
                                   plantilla,
                                   data,
                                   list_email,
                                   [],
                                   cuenta=CUENTAS_CORREOS[0][1]
                                   )
                    time.sleep(3)
                    cantidad_estudiantes_becados = BecaSolicitud.objects.values('id').filter(status=True, periodo_id=becasolicitud.periodo_id).exclude(becaaceptada=3).count()
        else:
            break

def generar_reporte_pendientes_cargar_requisitos(request, periodo):
    try:
        __author__ = 'Unemi'
        titulo = easyxf(
            'font: name Times New Roman, color-index black, bold on , height 350; alignment: horiz centre')
        titulo2 = easyxf(
            'font: name Times New Roman, color-index black, bold on , height 250; alignment: horiz centre')
        fuentecabecera = easyxf(
            'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
        fuentenormal = easyxf(
            'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
        fuentemoneda = easyxf(
            'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right',
            num_format_str=' "$" #,##0.00')
        fuentefecha = easyxf(
            'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz center',
            num_format_str='yyyy-mm-dd')
        fuentenumerodecimal = easyxf(
            'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right',
            num_format_str='#,##0.00')
        fuentenumeroentero = easyxf(
            'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right')

        font_style = XFStyle()
        font_style.font.bold = True
        font_style2 = XFStyle()
        font_style2.font.bold = False
        wb = Workbook(encoding='utf-8')
        ws = wb.add_sheet('Becados')
        response = HttpResponse(content_type="application/ms-excel")
        response['Content-Disposition'] = 'attachment; filename=becados_docupendiente_' + random.randint(1,
                                                                                                         10000).__str__() + '.xls'

        ws.write_merge(0, 0, 0, 12, 'UNIVERSIDAD ESTATAL DE MILAGRO', titulo)
        ws.write_merge(1, 1, 0, 12, 'DIRECCIÓN DE BIENESTAR UNIVERSITARIO', titulo2)
        ws.write_merge(2, 2, 0, 12, 'LISTADO DE BECADOS - DOCUMENTACIÓN PENDIENTE DE CARGAR', titulo2)

        row_num = 4
        columns = [
            (u"# SOLICITUD", 3000),
            (u"FECHA BECA", 3000),
            (u"MONTO BECA", 3000),
            (u"NOMBRES COMPLETOS", 10000),
            (u"IDENTIFICACIÓN", 5000),
            (u"FACULTAD", 5000),
            (u"CARRERA", 5000),
            (u"MODALIDAD", 5000),
            (u"NIVEL", 5000),
            (u"PROVINCIA", 5000),
            (u"CANTÓN", 5000),
            (u"PARROQUIA", 5000),
            (u"DIRECCIÓN", 15000),
            (u"REFERENCIA", 15000),
            (u"SECTOR", 15000),
            (u"# CASA", 5000),
            (u"# CORREO PERSONAL", 8000),
            (u"# CORREO UNEMI", 8000),
            (u"TELÉFONO", 5000),
            (u"CELULAR", 5000),
            (u"OPERADORA", 5000),
            (u"ESTADO CÉDULA", 5000),
            (u"ESTADO CUENTA BANCARIA", 5000)
        ]
        for col_num in range(len(columns)):
            ws.write(row_num, col_num, columns[col_num][0], fuentecabecera)
            ws.col(col_num).width = columns[col_num][1]

        row_num = 4

        becas = BecaAsignacion.objects.filter((Q(solicitud__inscripcion__persona__personadocumentopersonal__isnull=True) |
                                               Q(solicitud__inscripcion__persona__cuentabancariapersona__isnull=True)),
                                              status=True, solicitud__periodo=periodo,
                                              tipo=1, cargadocumento=True).order_by('-solicitud_id').distinct()

        for s in becas:
            row_num += 1
            p = Persona.objects.get(pk=s.solicitud.inscripcion.persona_id)
            documentos = p.documentos_personales()
            cuentabancaria = p.cuentabancaria()

            umat = s.solicitud.inscripcion.matricula_periodo_actual(periodo).first()
            ws.write(row_num, 0, str(s.solicitud.id), fuentenumeroentero)
            ws.write(row_num, 1, str(s.fecha_creacion)[:10], fuentefecha)
            ws.write(row_num, 2, s.montobeneficio, fuentemoneda)
            ws.write(row_num, 3, p.nombre_completo_inverso(), fuentenormal)
            ws.write(row_num, 4, p.identificacion(), fuentenormal)
            ws.write(row_num, 5, str(umat.nivel.coordinacion()) if umat else 'Retirado de Matricula del periodo actual', fuentenormal)
            ws.write(row_num, 6, str(umat.inscripcion.carrera) if umat else 'Retirado de Matricula del periodo actual', fuentenormal)
            ws.write(row_num, 7, str(umat.inscripcion.modalidad) if umat else 'Retirado de Matricula del periodo actual', fuentenormal)
            ws.write(row_num, 8, umat.nivelmalla.nombre if umat else 'Retirado de Matricula del periodo actual', fuentenormal)
            ws.write(row_num, 9, str(p.provincia) if p.provincia else '', fuentenormal)
            ws.write(row_num, 10, str(p.canton) if p.canton else '', fuentenormal)
            ws.write(row_num, 11, str(p.parroquia) if p.parroquia else '', fuentenormal)
            ws.write(row_num, 12, p.direccion_corta().upper(), fuentenormal)
            ws.write(row_num, 13, p.referencia.upper(), fuentenormal)
            ws.write(row_num, 14, p.sector.upper(), fuentenormal)
            ws.write(row_num, 15, p.num_direccion, fuentenormal)
            ws.write(row_num, 16, p.email, fuentenormal)
            ws.write(row_num, 17, p.emailinst, fuentenormal)
            ws.write(row_num, 18, p.telefono_conv, fuentenormal)
            ws.write(row_num, 19, p.telefono, fuentenormal)
            ws.write(row_num, 20, p.get_tipocelular_display() if p.tipocelular else '', fuentenormal)
            ws.write(row_num, 21, documentos.get_estadocedula_display() if documentos and documentos.estadocedula else 'SIN CARGAR', fuentenormal)
            ws.write(row_num, 22, cuentabancaria.get_estadorevision_display() if cuentabancaria else 'SIN CARGAR', fuentenormal)

        wb.save(response)
        return response
    except Exception as ex:
        raise NameError(str(ex))