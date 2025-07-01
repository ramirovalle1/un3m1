#!/usr/bin/env python
# -*- coding: utf-8 -*-
import requests
import io
import os
import sys
import json
import xlsxwriter
import xlwt
import openpyxl
import time
from xlwt import *

SITE_ROOT = os.path.dirname(os.path.realpath(__file__))
your_djangoproject_home = os.path.split(SITE_ROOT)[0]
sys.path.append(your_djangoproject_home)

from django.core.wsgi import get_wsgi_application

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "settings")
application = get_wsgi_application()
from sga.models import *
from sagest.models import *
from inno.models import CalendarioRecursoActividadAlumno, CalendarioRecursoActividad, \
    CalendarioRecursoActividadAlumnoMotificacion, MateriaAsignadaPlanificacionSedeVirtualExamen, \
    FechaPlanificacionSedeVirtualExamen, AulaPlanificacionSedeVirtualExamen, TurnoPlanificacionSedeVirtualExamen, \
    MatriculaSedeExamen
from bd.models import PeriodoCrontab
from django.db import transaction
from django.http import HttpResponse
from settings import DEBUG, HILOS_MAXIMOS
from wpush.models import SubscriptionInfomation
from webpush.models import SubscriptionInfo, PushInformation
from webpush.utils import _send_notification
from webpush import send_user_notification
from asgiref.sync import async_to_sync
from channels.layers import get_channel_layer
from ws.funciones import notificar_usuario_notificaciones
from sga.commonviews import traerNotificaciones
from sga.funciones import lista_gruposocioeconomico_beca
from matricula.models import CostoOptimoMalla, PeriodoMatricula
from sga.My_Model.MatriculaPregrado import My_MatriculacionPrimerNivelCarrera, My_ConfigMatriculacionPrimerNivel, \
    My_MatriculaPregrado
from matricula.funciones import get_tipo_matricula
from requests.auth import HTTPBasicAuth
from soap.consumer.senescyt import Titulos
from api.helpers.functions_helper import generate_qr_examen_final
from Moodle_Funciones import buscarQuiz, accesoQuizIndividual, estadoQuizIndividual


# ahora = datetime.now()
# manana = ahora + timedelta(days=1)
# pasadomanana = ahora + timedelta(days=2)
# fecha_hoy = ahora.date()
# hora_hoy = ahora.time()
# fecha_siguiente = manana.date()
# hora_siguiente = manana.time()
# fecha_pasado_siguiente = pasadomanana.date()
# hora_pasado_siguiente = pasadomanana.time()
# dia_semana_ahora = fecha_hoy.isoweekday()
# dia_semana_manana = fecha_siguiente.isoweekday()
# numerosemana = datetime.today().isocalendar()[1]


# aCaches = [f"noticia_panel_{ALUMNOS_GROUP_ID}",
#            f"banner_panel_{ALUMNOS_GROUP_ID}",
#            f"noticia_panel_{ALUMNOS_GROUP_ID}_admision",
#            f"banner_panel_{ALUMNOS_GROUP_ID}_admision",
#            f"noticia_panel_{ALUMNOS_GROUP_ID}_pregrado",
#            f"banner_panel_{ALUMNOS_GROUP_ID}_pregrado",
#            f"noticia_panel_{ALUMNOS_GROUP_ID}_posgrado",
#            f"banner_panel_{ALUMNOS_GROUP_ID}_posgrado"
#            ]
# for eCache in aCaches:
#     eNoticiaEnCache = cache.get(eCache)
#     if eNoticiaEnCache:
#         cache.delete(eCache)

# def segundos_a_segundos_minutos_y_horas(segundos):
#     horas = int(segundos / 60 / 60)
#     segundos -= horas*60*60
#     minutos = int(segundos/60)
#     segundos -= minutos*60
#     # return datetime(ahora.year, ahora.month, ahora.day, horas, minutos, segundos).strftime("%H:%M:%S")
#     return f"{int(horas):02d}:{int(minutos):02d}:{int(segundos):02d}"
#
#
# def notify_student_activities_pregrado(ePeriodo):
#     print(u"****************************************************************************************************")
#     print(f"Inicia proceso de notificar actividades de estudiantes a las {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} del periodo {ePeriodo}")
#     # calendario = calendar.Calendar()
#     eActividades = CalendarioRecursoActividad.objects.filter(Q(fechahorahasta__date__lte=fecha_siguiente) |
#                                                              Q(fechahorahasta__date__lte=fecha_pasado_siguiente),
#                                                              fechahoradesde__date__gte=fecha_hoy,
#                                                              materia__nivel__periodo=ePeriodo,
#                                                              materia__status=True, materia__cerrado=False,
#                                                              materia__nivel__cerrado=False,
#                                                              materia__inicio__lte=fecha_hoy,
#                                                              materia__fin__gte=fecha_hoy, status=True
#                                                              )
#     # if DEBUG:
#     eActividades = eActividades.filter(materia_id__in=MateriaAsignada.objects.values_list("materia_id", flat=True).filter(matricula_id__in=[472781, 464865]))
#     totalActividades = len(eActividades)
#     cotA = 0
#     for eActividad in eActividades:
#         cotA += 1
#         print(f"{totalActividades}/{cotA} --> {eActividad.__str__()}")
#         eMateria = eActividad.materia
#         segundos = (eActividad.fechahorahasta - datetime.now()).total_seconds()
#         SITE_URL_SIE = 'https://sgaestudiante.unemi.edu.ec'
#         if DEBUG:
#             SITE_URL_SIE = 'http://127.0.0.1:3000'
#         url = f"{SITE_URL_SIE}/alu_documentos"
#         titulo = f'Tiene una actividad: {eActividad.get_tipo_display()}'
#         convertido = segundos_a_segundos_minutos_y_horas(segundos)
#         tiempo = f"en {convertido}"
#         cuerpo = f'Su actividad {eActividad.get_tipo_display()} fue aperturada desde {eActividad.fechahoradesde.strftime("%d-%m-%Y")} se cierra {tiempo} ({eActividad.fechahorahasta.strftime("%d-%m-%Y")} a las {eActividad.fechahorahasta.strftime("%H:%M")})'
#         eAsignados = eMateria.asignados_a_esta_materia()
#         if DEBUG:
#             eAsignados = eAsignados.filter(matricula_id__in=[472781, 464865])
#         for eMateriaAsignada in eAsignados:
#             eInscripcion = eMateriaAsignada.matricula.inscripcion
#             ePersona = eInscripcion.persona
#             _eCalendarioRecursoActividadAlumno = CalendarioRecursoActividadAlumno.objects.filter(materiaasignada=eMateriaAsignada, recurso=eActividad, status=True)
#             if not _eCalendarioRecursoActividadAlumno.values("id").exists():
#                 eCalendarioRecursoActividadAlumno = CalendarioRecursoActividadAlumno(recurso=eActividad,
#                                                                                      materiaasignada=eMateriaAsignada)
#                 eCalendarioRecursoActividadAlumno.save()
#             else:
#                 eCalendarioRecursoActividadAlumno = _eCalendarioRecursoActividadAlumno[0]
#
#             ePerfilUsuario = PerfilUsuario.objects.filter(status=True, persona=ePersona, inscripcion=eInscripcion, visible=True)
#
#             #     else:
#             eCalendarioRecursoActividadAlumnoMotificaciones = CalendarioRecursoActividadAlumnoMotificacion.objects.filter(actividadalumno=eCalendarioRecursoActividadAlumno)
#             if not eCalendarioRecursoActividadAlumnoMotificaciones.values("id").exists():
#                 if ePerfilUsuario.values("id").exists():
#                     eNotificacion = Notificacion(titulo=titulo,
#                                                  cuerpo=cuerpo,
#                                                  destinatario=ePersona,
#                                                  perfil=ePerfilUsuario[0],
#                                                  url=url,
#                                                  prioridad=1,
#                                                  app_label='SIE',
#                                                  fecha_hora_visible=datetime.now() + timedelta(seconds=segundos),
#                                                  tipo=1,
#                                                  en_proceso=False,
#                                                  content_type=ContentType.objects.get_for_model(eActividad),
#                                                  object_id=eActividad.id,
#                                                  )
#                     eNotificacion.save()
#                     if segundos <= 86400:
#                         segundos_nueva = 86400
#                     elif segundos > 86400 and segundos <= 172800:
#                         segundos_nueva = 172800
#                     else:
#                         segundos_nueva = 259200
#                     eCalendarioRecursoActividadAlumnoMotificacion = CalendarioRecursoActividadAlumnoMotificacion(notificacion=eNotificacion,
#                                                                                                                  segundos=segundos_nueva,
#                                                                                                                  actividadalumno=eCalendarioRecursoActividadAlumno)
#                     eCalendarioRecursoActividadAlumnoMotificacion.save()
#             else:
#                 for eCalendarioRecursoActividadAlumnoMotificacion in eCalendarioRecursoActividadAlumnoMotificaciones:
#                     segundos_nueva = 0
#                     if segundos <= 86400:
#                         segundos_nueva = 86400
#                     elif segundos > 86400 and segundos <= 172800:
#                         segundos_nueva = 172800
#                     elif segundos > 172800 and segundos <= 259200:
#                         segundos_nueva = 259200
#
#                     bandera = segundos_nueva > 0 and segundos_nueva < eCalendarioRecursoActividadAlumnoMotificacion.segundos
#                     if DEBUG:
#                         bandera = True
#                     if bandera:
#                         if ePerfilUsuario.values("id").exists():
#                             eNotificacion = Notificacion(titulo=titulo,
#                                                          cuerpo=cuerpo,
#                                                          destinatario=ePersona,
#                                                          perfil=ePerfilUsuario[0],
#                                                          url=url,
#                                                          prioridad=1,
#                                                          app_label='SIE',
#                                                          fecha_hora_visible=datetime.now() + timedelta(seconds=segundos),
#                                                          tipo=1,
#                                                          en_proceso=False,
#                                                          content_type=ContentType.objects.get_for_model(eActividad),
#                                                          object_id=eActividad.id,
#                                                          )
#                             eNotificacion.save()
#                             # segundos_nueva = 86400
#                             eCalendarioRecursoActividadAlumnoMotificacion = CalendarioRecursoActividadAlumnoMotificacion(notificacion=eNotificacion,
#                                                                                                                          segundos=segundos_nueva,
#                                                                                                                          actividadalumno=eCalendarioRecursoActividadAlumno)
#                             eCalendarioRecursoActividadAlumnoMotificacion.save()
#
#
#                 subscriptions = ePersona.usuario.webpush_info.select_related("subscription")
#                 push_infos = SubscriptionInfomation.objects.filter(subscription_id__in=subscriptions.values_list('subscription__id', flat=True), app=2, status=True).select_related("subscription")
#                 for device in push_infos:
#                     payload = {
#                         "head": titulo,
#                         "body": cuerpo,
#                         "url": url,
#                         "action": "loadNotifications",
#                     }
#                     try:
#                         _send_notification(device.subscription, json.dumps(payload), ttl=500)
#                     except Exception as exep:
#                         print(f"Fallo de envio del push notification: {exep.__str__()}")
#                 channel_layer = get_channel_layer()
#                 async_to_sync(channel_layer.group_send)(f"{ePersona.usuario_id}", {"type": "cargar_ultimas_notificaciones",
#                                                                                    "message": "ping",
#                                                                                    "datos": {"persona_id": ePersona.id,
#                                                                                              "perfilusuario_id": ePerfilUsuario[0].id}}
#                                                         )
#     print(u"****************************************************************************************************")
#     print(u"****************************************************************************************************")
#     print(f"** Finaliza proceso de notificar actividades de estudiantes del periodo {ePeriodo}")


# def mis_materias(ePeriodo):
#     eExamenes = HorarioExamen.objects.filter(materia__nivel__periodo=ePeriodo)
#     eMaterias = Materia.objects.filter(nivel__periodo=ePeriodo, status=True, asignaturamalla__malla__carrera_id__in=[126, 135, 127, 133, 129, 131, 128, 132, 130, 134])
#     return eMaterias.exclude(pk__in=eExamenes.values_list("materia_id", flat=True)).order_by('asignaturamalla__malla', 'asignaturamalla__nivelmalla__orden', 'paralelomateria')
#
#
# import datetime
#
# ePeriodo = Periodo.objects.get(pk=126)
#
# turnos_nocturnbos = [
#     [datetime.time(17, 0, 0), datetime.time(18, 59, 59)],
#     [datetime.time(19, 0, 0), datetime.time(20, 59, 59)],
#     [datetime.time(21, 0, 0), datetime.time(22, 59, 59)],
# ]
# fechas = [
#     datetime.datetime(2022, 7, 18),
#     datetime.datetime(2022, 7, 19),
#     datetime.datetime(2022, 7, 20),
#     datetime.datetime(2022, 7, 21),
#     datetime.datetime(2022, 7, 22),
#     datetime.datetime(2022, 7, 25),
#     datetime.datetime(2022, 7, 26),
#     datetime.datetime(2022, 7, 27),
#     datetime.datetime(2022, 7, 28),
#     datetime.datetime(2022, 7, 29),
# ]
# total = len(turnos_nocturnbos) * len(fechas)
# detallemodelo_id=33
# aula_id=218
# contador = 0
# isBreak = True
# while isBreak:
#     contador = 0
#     eMaterias = mis_materias(ePeriodo)
#     if not eMaterias.values("id").exists():
#         contador = total
#         break
#     for turno in turnos_nocturnbos:
#         eMaterias = mis_materias(ePeriodo)
#         if not eMaterias.values("id").exists():
#             contador = total
#             break
#         for fecha in fechas:
#             contador += 1
#             eMaterias = mis_materias(ePeriodo)
#             if not eMaterias.values("id").exists():
#                 contador = total
#                 break
#             if eMaterias.values("id").exists():
#                 eMateria = eMaterias.first()
#                 eAsignados = eMateria.materiaasignada_set.select_related().filter(status=True, retiramateria=False)
#                 cantidadalumnos = len(eAsignados)
#                 eProfesorMaterias = eMateria.profesormateria_set.filter(Q(status=True), Q(activo=True),
#                                                                         (
#                                                                                 Q(tipoprofesor__id=TIPO_DOCENTE_TEORIA) |
#                                                                                 Q(tipoprofesor__id=TIPO_DOCENTE_FIRMA) |
#                                                                                 Q(tipoprofesor__id__in=[9, 11, 12, 10, 7, 14])
#                                                                         )
#                                                                         )
#                 eProfesorMateria =None
#                 if eProfesorMaterias.values("id").exists():
#                     eProfesorMateria = eProfesorMaterias.first()
#                 maximo = variable_valor('MAXIMO_ESTUDIANTES_EXAMEN')
#                 puedo_insertar = null_to_numeric(HorarioExamenDetalle.objects.filter(status=True, horarioexamen__fecha=fecha, horainicio=turno[0], horafin=turno[1], horarioexamen__detallemodelo_id=detallemodelo_id).aggregate(total=Sum('cantalumnos')).get('total')) + int(cantidadalumnos) <= maximo and int(cantidadalumnos) <= maximo
#                 if puedo_insertar:
#                     if not HorarioExamen.objects.filter(materia=eMateria).exists():
#                         eHorarioExamen = HorarioExamen(materia=eMateria,
#                                                        fecha=fecha,turno_id=1,
#                                                        detallemodelo_id=detallemodelo_id,
#                                                        )
#                         eHorarioExamen.save()
#                         contador -= 1
#                         print(u"Inserta horario de materia %s"%eHorarioExamen)
#                     else:
#                         eHorarioExamen = HorarioExamen.objects.filter(materia=eMateria).first()
#                     if not HorarioExamenDetalle.objects.filter(horarioexamen=eHorarioExamen).exists():
#                         eHorarioExamenDetalle = HorarioExamenDetalle(horarioexamen=eHorarioExamen,
#                                                                      horainicio=turno[0],
#                                                                      horafin=turno[1],
#                                                                      cantalumnos=cantidadalumnos,
#                                                                      aula_id=aula_id,
#                                                                      profesormateria=eProfesorMateria
#                                                                      )
#                         eHorarioExamenDetalle.save()
#                         print(u"Inserta detalle horario de materia %s" % eHorarioExamenDetalle)
#                     else:
#                         eHorarioExamenDetalle = HorarioExamenDetalle.objects.filter(horarioexamen=eHorarioExamen).first()
#                     print(u"CONTADOR:   %s"%contador)
#                     eliminar = HorarioExamenDetalleAlumno.objects.filter(horarioexamendetalle=eHorarioExamenDetalle)
#                     eliminar.delete()
#                     print(u"Inserta horario con %s alumnos" % cantidadalumnos)
#                     for eAsignado in eAsignados:
#                         eHorarioExamenDetalleAlumno = HorarioExamenDetalleAlumno(horarioexamendetalle=eHorarioExamenDetalle,
#                                                                                  materiaasignada=eAsignado)
#                         eHorarioExamenDetalleAlumno.save()
#     if contador == total:
#         print(u"Entra al break")
#         isBreak = False

# eMateriaAsignadas = MateriaAsignada.objects.filter(matricula__nivel__periodo=ePeriodo)
# eMateriaAsignadas.update(asistenciafinal=100, sinasistencia=True)
# total_ma = len(eMateriaAsignadas)
# print(f"Se actualiza estado sin asistencia a materias asignadas un total de: {total_ma} del periodo {ePeriodo.__str__()}")
# eMaterias = Materia.objects.filter(nivel__periodo=ePeriodo)
# if DEBUG:
#     eMaterias = eMaterias[0:10]
# total_m = len(eMaterias)
# print(f"Se va a proceder actualizar modelo evaluativo un total de: {total_m} del periodo {ePeriodo.__str__()}")
# contador = 0
# modeloevaluativo_id = 24
# for eMateria in eMaterias:
#     contador += 1
#     eMateria.modeloevaluativo_id = modeloevaluativo_id
#     eMateria.save()
#     evaluaciones = EvaluacionGenerica.objects.filter(materiaasignada__materia=eMateria)
#     evaluaciones.delete()
#     for maa in eMateria.asignados_a_esta_materia():
#         maa.evaluacion()
#         maa.notafinal = 0
#         maa.save(actualiza=False)
#     if eMateria.cronogramaevaluacionmodelo_set.exists():
#         cronograma = eMateria.cronogramaevaluacionmodelo_set.all()[0]
#         cronograma.materias.remove(eMateria)
#
#     print(f"Materia ({contador}/{total_m}) ---> {eMateria.__str__()}")
# notify_student_activities_pregrado(ePeriodo)
# eLeccionGrupos = LeccionGrupo.objects.filter(status=False)
# eLeccionGrupos.delete()
# eMateriaAsignadas = MateriaAsignada.objects.filter(matricula__nivel__periodo=ePeriodo)
# eMaterias = Materia.objects.filter(pk__in=eMateriaAsignadas.values_list("materia__id", flat=True))
# totalMaterias = len(eMaterias)
# print(f"********************** TOTAL DE MATERIAS {totalMaterias} del periodo {ePeriodo}")
# contadorMateria = 0
# for eMateria in eMaterias:
#     contadorMateria += 1
#     print(f"********************** {contadorMateria}/{totalMaterias} de la {eMateria} ... INICIA")
#     crear_editar_calendario_actividades_pregrado(eMateria)
#     print(f"********************** {contadorMateria}/{totalMaterias} de la {eMateria} ... FINALIZA")


# ePeriodos = Periodo.objects.filter(pk__in=[85, 119])
# ePeriodos = Periodo.objects.filter(pk__in=[124,123,122,121,113,112,111,110])
# ePeriodos = Periodo.objects.filter(pk__in=[99,97,96,95,94,90,89,83,82,80])
# ePeriodos = Periodo.objects.filter(pk__in=[77,76,75,74,73,72,71,70,69,68,67,66,65,64,63,62,61,60,59,58])
# ePeriodos = Periodo.objects.filter(pk__in=[57,56,55,54,53,52,51,50,49,48,47,46,45,44,43,42,41,40,39,38,37,36,35,34,33,32,31,30,29,28,27,26,25,24,23,22,21,20,19,18,17,16,15,14,11,10,9,7,6,5])
# eliminar_lecciongrupos(ePeriodos)
# cambiar_fecha_rubros(ePeriodo, datetime(2022, 5, 30, 0, 0, 0).date())
# reajuste_horario_clase_seguida_pregrado(ePeriodo.id)


# ePeriodo = Periodo.objects.get(pk=126)
# fecha_hoy = date(2022, 7, 18)
# def eliminar_lecciones_no_aperturada(periodo, fecha_hoy):
#     print(u"****************************************************************************************************")
#     print(f"Inicia proceso de eliminar lecciones no aperturadas a las {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} del periodo {periodo}")
#     eProfesores = Profesor.objects.filter(profesormateria__materia__nivel__periodo=periodo, profesormateria__activo=True).distinct()
#     total_profesores = len(eProfesores)
#     contP = 1
#     for eProfesor in eProfesores:
#         print(f"***** ({contP}/{total_profesores}) -> Profesor: {eProfesor.__str__()}")
#         contP += 1
#         eLeccionGrupos = LeccionGrupo.objects.filter(status=False, fecha__lte=fecha_hoy, profesor=eProfesor, lecciones__clase__profesor=eProfesor, lecciones__fecha__lte=fecha_hoy)
#         with transaction.atomic():
#             try:
#                 for eLeccionGrupo in eLeccionGrupos:
#                     eLecciones = eLeccionGrupo.lecciones.all()
#                     if eLecciones.values("id").filter(status=True).exists():
#                         for eLeccion in eLecciones:
#                             for eAsistenciaLeccion in eLeccion.asistencialeccion_set.all():
#                                 eAsistenciaLeccion.status = True
#                                 eAsistenciaLeccion.save()
#                                 eMateriaAsignada = eAsistenciaLeccion.materiaasignada
#                                 eMateriaAsignada.save(actualiza=True)
#                             eLeccion.status = True
#                             eLeccion.save()
#                         eLeccionGrupo.status = True
#                         eLeccionGrupo.save()
#                     else:
#                         eLeccionGrupo.delete()
#             except Exception as ex:
#                 transaction.set_rollback(True)
#     print(u"****************************************************************************************************")
#     print(u"****************************************************************************************************")
#     print(f"** Finaliza proceso de eliminar lecciones no aperturadas del periodo {periodo}")


def eliminar_lecciones(periodo, fecha_hoy):
    numerosemanaactual = datetime.today().isocalendar()[1]
    print(u"****************************************************************************************************")
    print(
        f"Inicia proceso de eliminar lecciones aperturadas a las {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} del periodo {periodo}")
    eProfesores = Profesor.objects.filter(profesormateria__materia__nivel__periodo=periodo,
                                          profesormateria__activo=True).distinct()
    total_profesores = len(eProfesores)
    contP = 1
    for eProfesor in eProfesores:
        print(f"***** ({contP}/{total_profesores}) -> Profesor: {eProfesor.__str__()}")
        contP += 1
        eLeccionGrupos = LeccionGrupo.objects.filter(fecha=fecha_hoy, profesor=eProfesor,
                                                     lecciones__clase__profesor=eProfesor, lecciones__fecha=fecha_hoy)
        with transaction.atomic():
            try:
                for eLeccionGrupo in eLeccionGrupos:
                    eLecciones = eLeccionGrupo.lecciones.all()
                    tieneferiado = False
                    if eLecciones.values("id").exists():
                        for eLeccion in eLecciones:
                            eClase = eLeccion.clase
                            fechacompara = eClase.compararfecha(numerosemanaactual)
                            tieneferiado = periodo.es_feriado(fechacompara, eClase.materia)
                            if tieneferiado:
                                break
                    else:
                        eLeccionGrupo.delete()
                    if tieneferiado:
                        eLeccionGrupo.delete()
            except Exception as ex:
                transaction.set_rollback(True)
    print(u"****************************************************************************************************")
    print(u"****************************************************************************************************")
    print(f"** Finaliza proceso de eliminar lecciones no aperturadas del periodo {periodo}")


def asistencia_integracion_curricular_por_plananalitico_alumnos():
    print(f"***** INICIA PROCESO DE ASISTENTE DE INTEGRACIÓN CURRICULAR")
    # fechas = [
    #     (datetime(2022, 6, 13, 0, 0, 0)).date(),
    #     (datetime(2022, 6, 14, 0, 0, 0)).date(),
    #     (datetime(2022, 6, 15, 0, 0, 0)).date(),
    #     (datetime(2022, 6, 16, 0, 0, 0)).date(),
    #     (datetime(2022, 6, 17, 0, 0, 0)).date(),
    #     (datetime(2022, 6, 20, 0, 0, 0)).date(),
    #     (datetime(2022, 6, 21, 0, 0, 0)).date(),
    #     (datetime(2022, 6, 22, 0, 0, 0)).date(),
    #     (datetime(2022, 6, 23, 0, 0, 0)).date(),
    #     (datetime(2022, 6, 24, 0, 0, 0)).date(),
    #     (datetime(2022, 6, 27, 0, 0, 0)).date(),
    #     (datetime(2022, 6, 28, 0, 0, 0)).date(),
    #     (datetime(2022, 6, 29, 0, 0, 0)).date(),
    #     (datetime(2022, 6, 30, 0, 0, 0)).date(),
    # ]
    fechas = [
        (datetime(2022, 6, 13, 0, 0, 0)).date(),
        (datetime(2022, 6, 14, 0, 0, 0)).date(),
        (datetime(2022, 6, 15, 0, 0, 0)).date(),
        (datetime(2022, 6, 16, 0, 0, 0)).date(),
        (datetime(2022, 6, 17, 0, 0, 0)).date(),
        (datetime(2022, 6, 20, 0, 0, 0)).date(),
        (datetime(2022, 6, 21, 0, 0, 0)).date(),
        (datetime(2022, 6, 22, 0, 0, 0)).date(),
        (datetime(2022, 6, 23, 0, 0, 0)).date(),
        (datetime(2022, 6, 24, 0, 0, 0)).date(),
        (datetime(2022, 6, 27, 0, 0, 0)).date(),
        (datetime(2022, 6, 28, 0, 0, 0)).date(),
        (datetime(2022, 6, 29, 0, 0, 0)).date(),
        (datetime(2022, 6, 30, 0, 0, 0)).date(),
    ]
    ePeriodo = Periodo.objects.get(pk=126)
    ePeriodoAcademia = ePeriodo.get_periodoacademia()

    for fecha in fechas:
        ids_p = []
        dia = (fecha.weekday() + 1)

        profesores = Profesor.objects.filter(profesormateria__materia__nivel__periodo=ePeriodo,
                                             profesormateria__activo=True,
                                             profesormateria__materia__nivel_id__in=[717]).distinct()

        # profesores = Profesor.objects.filter(profesormateria__materia__nivel__periodo=ePeriodo, profesormateria__activo=True).distinct()
        # profesores = Profesor.objects.filter(pk=903).distinct()
        total_profesores = profesores.count()
        print("*** FECHA A PROCESAR: " + fecha.__str__() + "\r")
        contP = 1
        for profesor in profesores:
            print(f"***** ({contP}/{total_profesores}) -> Profesor: {profesor.__str__()}")
            # clases = Clase.objects.filter(activo=True, inicio__lte=fecha, fin__gte=fecha, dia=dia, status=True, materia__nivel__periodo_id=119, profesor_id=profesor.id).exclude()
            clases = Clase.objects.filter(activo=True, inicio__lte=fecha, fin__gte=fecha, dia=dia, status=True,
                                          materia__nivel__periodo=ePeriodo,
                                          materia__nivel__periodo__visible=True,
                                          materia__nivel__periodo__visiblehorario=True,
                                          materia__fechafinasistencias__gte=fecha,
                                          materia__profesormateria__profesor_id=profesor.id,
                                          profesor_id=profesor.id,
                                          materia__nivel_id__in=[717])
            # clases = clases.filter(turno__comienza__gte=time(18, 0, 0), turno__termina__lte=time(21, 59, 00))
            total_clases = clases.count()
            contC = 1
            for cl in clases:
                print(f"******** ({contC}/{total_clases}) -> Clase: {cl.__str__()}")
                lecciongrupo = None
                if LeccionGrupo.objects.values("id").filter(profesor=profesor, turno=cl.turno, fecha=fecha).exists():
                    lecciongrupo = LeccionGrupo.objects.get(profesor=profesor, turno=cl.turno, fecha=fecha)
                    lecciongrupo.abierta = False
                    lecciongrupo.save()
                else:
                    lecciongrupo = LeccionGrupo(profesor=profesor,
                                                turno=cl.turno,
                                                aula=cl.aula,
                                                dia=cl.dia,
                                                fecha=fecha,
                                                horaentrada=cl.turno.comienza,
                                                horasalida=cl.turno.termina,
                                                abierta=False,
                                                automatica=True,
                                                contenido='REGISTRO MASIVO 2022 - AUTORIZADO POR DIRECTOR TICS',
                                                observaciones='REGISTRO MASIVO 2022 - AUTORIZADO POR DIRECTOR TICS')
                    lecciongrupo.save()
                if lecciongrupo:
                    leccion = None
                    if Leccion.objects.values("id").filter(clase=cl, fecha=fecha).exists():
                        leccion = Leccion.objects.get(clase=cl, fecha=fecha)
                        leccion.abierta = False
                        leccion.save()
                    else:
                        leccion = Leccion(clase=cl,
                                          fecha=fecha,
                                          horaentrada=cl.turno.comienza,
                                          horasalida=cl.turno.termina,
                                          abierta=False,
                                          contenido=lecciongrupo.contenido,
                                          observaciones=lecciongrupo.observaciones)
                        leccion.save()
                    if leccion:
                        if not lecciongrupo.lecciones.values("id").filter(pk=leccion.id).exists():
                            lecciongrupo.lecciones.add(leccion)

                        asignados = None
                        # SE FILTRA SI LA MATERIA TIENE TIPO DE PROFESOR PRACTICA Y LA CLASE TAMBIEN
                        # 1 => CLASE PRESENCIAL
                        # 2 => CLASE VIRTUAL SINCRÓNICA
                        # 8 => CLASE REFUERZO SINCRÓNICA
                        if cl.tipoprofesor_id == 2 and cl.tipohorario in [1, 2, 8]:
                            if cl.grupoprofesor:
                                if cl.grupoprofesor.paralelopractica:
                                    # grupoprofesor_id = clase.grupoprofesor.id
                                    if cl.grupoprofesor.listado_inscritos_grupos_practicas().exists():
                                        listado_alumnos_practica = cl.grupoprofesor.listado_inscritos_grupos_practicas()
                                        if ePeriodoAcademia.valida_asistencia_pago:
                                            asignados = MateriaAsignada.objects.filter(
                                                pk__in=listado_alumnos_practica.values_list('materiaasignada_id',
                                                                                            flat=True),
                                                matricula__estado_matricula__in=[2, 3]).distinct()
                                        else:
                                            asignados = MateriaAsignada.objects.filter(
                                                pk__in=listado_alumnos_practica.values_list('materiaasignada_id',
                                                                                            flat=True)).distinct()
                                    else:
                                        print(u"Clase no creada, de tipo práctica no tiene alumnos. %s" % cl)
                                else:
                                    print(u"Clase no creada, de tipo práctica no tiene paralelos. %s" % cl)
                            else:
                                print(u"Clase no creada, de tipo práctica no tiene grupos. %s" % cl)
                        else:
                            asignados = cl.materia.asignados_a_esta_materia()

                        total_asistencias = 0
                        if AsistenciaLeccion.objects.values("id").filter(leccion=leccion).exists():
                            total_asistencias = AsistenciaLeccion.objects.filter(leccion=leccion).count()
                            for asis in AsistenciaLeccion.objects.filter(leccion=leccion):
                                if not asis.asistio:
                                    asis.asistio = True
                                    asis.save()
                                    mateasig = asis.materiaasignada
                                    mateasig.save(actualiza=True)
                                    # mateasig.actualiza_estado()
                        else:
                            total_asistencias = asignados.count()
                            for materiaasignada in asignados:
                                if not AsistenciaLeccion.objects.values("id").filter(leccion=leccion,
                                                                                     materiaasignada=materiaasignada).exists():
                                    asistencialeccion = AsistenciaLeccion(leccion=leccion,
                                                                          materiaasignada=materiaasignada,
                                                                          asistio=True)
                                    asistencialeccion.save()
                                    materiaasignada.save(actualiza=True)
                                    # materiaasignada.actualiza_estado()
                                # guardar temas de silabo
                        lecciongrupo.save()
                        print(f"*********** (Total Asistencia: {total_asistencias})")
                # print(cl)
                contC += 1

            contP += 1


def reajuste_horario_examen(ePeriodo):
    eMaterias = Materia.objects.filter(status=True, nivel__periodo=ePeriodo)
    totalMaterias = eMaterias.count()
    print(f"*->Total de materias {totalMaterias}")
    contador = 0
    for eMateria in eMaterias:
        contador += 1
        eHorarioExamenes = HorarioExamen.objects.filter(status=True, materia=eMateria)
        print(f"******************************************************************")
        print(f"******-> ({contador} / {totalMaterias}) Horario de examen de la materia: {eMateria.__str__()}")
        eAlumnos = MateriaAsignada.objects.filter(status=True, retiramateria=False, materia=eMateria,
                                                  matricula__retiradomatricula=False, matricula__status=True).order_by(
            'matricula__inscripcion__persona__apellido1', 'matricula__inscripcion__persona__apellido2',
            'matricula__inscripcion__persona__nombres')
        print(f"******->Total de matriculados ({eAlumnos.count()}) en la materia: {eMateria.__str__()}")
        for eHorarioExamen in eHorarioExamenes:
            eHorarioExamenDetalles = HorarioExamenDetalle.objects.filter(status=True, horarioexamen=eHorarioExamen)
            tieneAlumnosAsignados = True
            for eHorarioExamenDetalle in eHorarioExamenDetalles:
                eAsignados = HorarioExamenDetalleAlumno.objects.filter(horarioexamendetalle=eHorarioExamenDetalle,
                                                                       status=True, materiaasignada__in=eAlumnos)
                cantalumnos = eHorarioExamenDetalle.cantalumnos
                if cantalumnos > 0 and eAsignados.count() == 0:
                    tieneAlumnosAsignados = False
            if not tieneAlumnosAsignados:
                print(f"************->No tiene alumnos asignados al horario")
                eAsignados = HorarioExamenDetalleAlumno.objects.filter(horarioexamendetalle__in=eHorarioExamenDetalles,
                                                                       status=True, materiaasignada__in=eAlumnos)
                if eAsignados.values("id").exists():
                    eAsignados.delete()
                for eHorario in eHorarioExamenDetalles:
                    print(f"************************->Inicia proceso de asignar en el horario {eHorario.__str__()}")

                    eAsignados = HorarioExamenDetalleAlumno.objects.filter(
                        horarioexamendetalle__in=eHorarioExamenDetalles, status=True, materiaasignada__in=eAlumnos)
                    cantidad = eHorario.cantalumnos
                    inicial = 0
                    ultimo = cantidad
                    for eAlumno in eAlumnos.exclude(pk__in=eAsignados.values_list('materiaasignada_id', flat=True))[
                                   inicial:ultimo]:
                        eHorarioExamenDetalleAlumno = HorarioExamenDetalleAlumno(materiaasignada=eAlumno,
                                                                                 horarioexamendetalle=eHorario)
                        eHorarioExamenDetalleAlumno.save()
                    print(f"************************->Finalizó proceso de asignar en el horario {eHorario.__str__()}")


def eliminar_lecciones_dia_especifico():
    print(f"***** INICIA PROCESO DE ELIMINAR LECCIONES DEL DÍA 2022-07-18")
    ids_lecciongrupos = [893641, 893646, 893702, 893678, 893635, 893653, 893745, 893762, 893858, 893858, 893628, 893659,
                         893777, 893778, 893673, 893676, 893687, 893681, 893740, 893683, 893739, 893760, 893815, 893829,
                         893789, 893808, 893748, 893680, 893741, 893637, 893651, 893677, 893688, 893736, 893623, 893663,
                         893701, 893618, 893668, 893765, 893624, 893662, 893695, 893674, 893690, 893625, 893661, 893856,
                         893857, 893629, 893658, 893694, 893779, 893814, 893832, 893632, 893655, 893803, 893784, 893785,
                         893786, 893787, 893672, 893709, 893730, 893763, 893792, 893806, 893797, 893630, 893657, 893642,
                         893647, 893692, 893679, 893686, 893746, 893820, 893821, 893822, 893780, 893813, 893638, 893650,
                         893791, 893807, 893827, 893783, 893810, 893781, 893812, 893825, 893826, 893626, 893743, 893750,
                         893768, 893729, 893752, 893774, 893620, 893666, 893696, 893643, 893645, 893707, 893742, 893753,
                         893772, 893617, 893669, 893703, 893616, 893670, 893716, 893744, 893751, 893775, 893747, 893761,
                         893639, 893649, 893698, 893798, 893799, 893800, 893801, 893824, 893782, 893811, 893831, 893823,
                         893828, 893843, 893636, 893652, 893706, 893622, 893664, 893699, 893788, 893809, 893830, 893621,
                         893665, 893712, 893731, 893728, 893757, 893802, 93805, 893644, 893691, 893738, 93756, 893804,
                         893759, 893769, 893675, 893689, 893671, 893682, 893764]
    eLeccionGrupos = LeccionGrupo.objects.filter(status=True, lecciones__clase__materia__status=True,
                                                 lecciones__clase__materia__planificacionclasesilabo_materia__tipoplanificacion_id=815,
                                                 fecha=(datetime(2022, 7, 18, 0, 0, 0)).date())
    total = eLeccionGrupos.count()
    contP = 0
    for eLeccionGrupo in eLeccionGrupos:
        contP += 1
        print(f"***** ({contP}/{total}) -> LecciónGrupo: {eLeccionGrupo.__str__()}")
        eLecciones = eLeccionGrupo.mis_leciones()
        for eLeccion in eLecciones:
            print(f"***** Elimna Lección: {eLeccion.__str__()}")
            eLeccion.delete()
        print(f"***** Elimna LecciónGrupo: {eLeccionGrupo.__str__()}")
        eLeccionGrupo.delete()


def actualizar_asistencia(ePeriodo):
    print(f"***** INICIA PROCESO DE ACTUALIZAR PORCENJATE DE ASISTENCIA")
    eMatriculas = Matricula.objects.filter(status=True, bloqueomatricula=False, retiradomatricula=False,
                                           inscripcion__modalidad_id__in=[1, 2], nivel__periodo=ePeriodo)
    total = eMatriculas.count()
    contP = 0
    for eMatricula in eMatriculas:
        contP += 1
        print(f"***** ({contP}/{total}) -> Matricula: {eMatricula.__str__()}")
        eMateriaAsignadas = MateriaAsignada.objects.filter(status=True, matricula=eMatricula, retiramateria=False)
        for eMateriaAsignada in eMateriaAsignadas:
            print(f"*************************** Materia: {eMateriaAsignada.materia.__str__()}")
            eMateriaAsignada.save(actualiza=True)


def descarga_resultado_encuesta(id, nombre_archivo):
    response = HttpResponse(content_type="application/ms-excel")
    response['Content-Disposition'] = f'attachment; filename={nombre_archivo}.xls'
    style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
    style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
    style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
    title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
    style1 = easyxf(num_format_str='D-MMM-YY')
    font_style = XFStyle()
    font_style.font.bold = True
    font_style2 = XFStyle()
    font_style2.font.bold = False
    wb = xlwt.Workbook()
    ws = wb.add_sheet('Sheetname')
    estilo = xlwt.easyxf(
        'font: height 350, name Arial, colour_index black, bold on, italic on; align: wrap on, vert centre, horiz center;')
    ws.write_merge(0, 0, 0, 9, 'UNIVERSIDAD ESTATAL DE MILAGRO', estilo)
    output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media', 'reportes', 'encuestas'))
    try:
        os.stat(output_folder)
    except:
        os.mkdir(output_folder)
    nombre = nombre_archivo + "_" + datetime.now().strftime('%Y%m%d_%H%M%S') + ".xls"
    filename = os.path.join(output_folder, nombre)
    encuesta = EncuestaGrupoEstudiantes.objects.get(id=id)
    ePersona = Persona.objects.get(pk=10730)
    usernotify = ePersona.usuario
    eNotificacion = Notificacion(cuerpo='Generación de reporte de excel en progreso',
                                 titulo=f'En proceso encuesta ID: {id}',
                                 destinatario=ePersona,
                                 url='',
                                 prioridad=1,
                                 app_label='SGA',
                                 fecha_hora_visible=datetime.now() + timedelta(days=1),
                                 tipo=2,
                                 en_proceso=True)
    eNotificacion.save()
    preguntas = encuesta.preguntaencuestagrupoestudiantes_set.filter(status=True).order_by('orden')
    columns = [(u"Nº.", 2000),
               (u"ID", 2000),
               (u"RESPONDIO", 3000),
               (u"CÉDULA", 3000),
               (u"ENCUESTADO", 9000),
               ]

    if encuesta.tipoperfil == 1:  # ALUMNO
        columns.append((u'CARRERA', 9000), )
        # solo para encuesta con y sin discapacidad
        if encuesta.id == 20 or encuesta.id == 21 or encuesta.id == 22 or encuesta.id == 23:  # encuestas
            columns.append((u'TIENE DISCAPACIDAD', 3000), )
        # fin

    if encuesta.tipoperfil == 2:  # DOCENTE
        columns.append((u'Tipo de relación laboral', 9000), )
        columns.append((u'Tiempo de dedicación', 9000), )
        columns.append((u'Si es docente titular a qué categoría académica pertenece', 9000), )
        # columns.append((u'MODALIDAD CONTRATACIÓN', 9000), )
        columns.append((u'A qué Facultad pertenece', 9000), )
        columns.append((u'Las carreras en las que imparte docencia actualmente ¿De qué modalidad son? ', 9000), )
        # columns.append((u'MODALIDAD DE LA CARRERA QUE DESEARÍA TRABAJAR EN EL SEMESTRE 1S 2022', 9000), )

    if encuesta.tipoperfil == 3:  # ADMINISTRATIVO
        columns.append((u'Tipo de relación laboral', 9000), )
        columns.append((u'Denominación del puesto', 9000), )

    for x in preguntas:
        columns.append((str(x.orden) + ") " + x.descripcion, 6000), )
        if x.tipo == 1:
            if not x.esta_vacia():
                columns.append((str(x.orden) + ") " + x.observacionporno, 6000), )
    row_num = 3
    for col_num in range(len(columns)):
        ws.write(row_num, col_num, columns[col_num][0], font_style)
        ws.col(col_num).width = columns[col_num][1]
    row_num = 4
    i = 0
    datos = []
    if encuesta.tipoperfil == 1:  # ALUMNO
        datos = encuesta.inscripcionencuestagrupoestudiantes_set.filter(status=True).order_by(
            'inscripcion__persona__apellido1')
    if encuesta.tipoperfil == 2:  # DOCENTE
        datos = encuesta.inscripcionencuestagrupoestudiantes_set.filter(status=True).order_by(
            'profesor__persona__apellido1')
    if encuesta.tipoperfil == 3:  # ADMINISTRATIVO
        datos = encuesta.inscripcionencuestagrupoestudiantes_set.filter(status=True).order_by(
            'administrativo__persona__apellido1')
    cout_register = len(datos)
    register_start = 0
    limit = 0
    for dato in datos:
        try:
            row_num += limit
            i += 1
            limit = 0
            ws.write(row_num, 0, i, font_style2)

            if encuesta.tipoperfil == 1:
                ws.write(row_num, 1, dato.inscripcion_id, font_style2)
                ws.write(row_num, 2, dato.respondio, font_style2)
                ws.write(row_num, 3, dato.inscripcion.persona.documento(), font_style2)
                ws.write(row_num, 4, dato.inscripcion.persona.nombre_completo_inverso(), font_style2)
            if encuesta.tipoperfil == 2:
                ws.write(row_num, 1, dato.profesor_id, font_style2)
                ws.write(row_num, 2, dato.respondio, font_style2)
                ws.write(row_num, 3, dato.profesor.persona.documento(), font_style2)
                ws.write(row_num, 4, dato.profesor.persona.nombre_completo_inverso(), font_style2)
            if encuesta.tipoperfil == 3:
                ws.write(row_num, 1, dato.administrativo_id, font_style2)
                ws.write(row_num, 2, dato.respondio, font_style2)
                ws.write(row_num, 3, dato.administrativo.persona.documento(), font_style2)
                ws.write(row_num, 4, dato.administrativo.persona.nombre_completo_inverso(), font_style2)
            c = 5
            if encuesta.tipoperfil == 1:
                ws.write(row_num, c, dato.inscripcion.carrera.__str__(),
                         font_style2) if not dato.inscripcion.carrera == None else ' '
                c += 1

            # solo para encuesta con y sin discapacidad
            if encuesta.id == 20 or encuesta.id == 21 or encuesta.id == 22 or encuesta.id == 23:  # encuesta
                if dato.inscripcion.persona.mi_perfil().tienediscapacidad:
                    discapacidad = 'SI'
                else:
                    discapacidad = 'NO'

                ws.write(row_num, c, discapacidad, font_style2)
                c += 1
            # fin

            if encuesta.tipoperfil == 2:
                dt = ProfesorDistributivoHoras.objects.filter(status=True, periodo=126,
                                                              profesor_id=dato.profesor.id).first()
                ws.write(row_num, c, dt.nivelcategoria.nombre if dt is not None else '', font_style2)
                c += 1
                ws.write(row_num, c, dt.dedicacion.nombre if dt is not None else '', font_style2)
                c += 1
                ws.write(row_num, c, dt.categoria.nombre if dt is not None and dt.nivelcategoria.id == 1 else '',
                         font_style2)
                c += 1
                ws.write(row_num, c, dt.coordinacion.nombre if dt is not None and dt.coordinacion is not None else '',
                         font_style2)
                c += 1
                w = 0
                for m in dato.profesor.mis_materias(126).values_list('materia__nivel__modalidad__nombre',
                                                                     flat=True).distinct(
                        'materia__nivel__modalidad__nombre'):
                    ws.write(row_num + w, c, str(m), font_style2)
                    w += 1
                if limit < w and w > 0:
                    limit = w - 1

                c += 1
            if encuesta.tipoperfil == 3:
                eDistributivoPersonas = DistributivoPersona.objects.filter(persona=dato.administrativo.persona,
                                                                           status=True, regimenlaboral_id=2,
                                                                           estadopuesto_id=1)
                eDistributivoPersona = None
                if eDistributivoPersonas.values("id").exists():
                    eDistributivoPersona = eDistributivoPersonas.first()
                ws.write(row_num, c,
                         eDistributivoPersona.regimenlaboral.descripcion if eDistributivoPersona is not None else '',
                         font_style2)
                c += 1
                ws.write(row_num, c,
                         eDistributivoPersona.denominacionpuesto.descripcion if eDistributivoPersona is not None else '',
                         font_style2)
                c += 1
            if dato.respondio:
                for x in preguntas:
                    respuesta = None
                    if x.tipo == 1:
                        respuesta = RespuestaPreguntaEncuestaGrupoEstudiantes.objects.filter(status=True, pregunta=x,
                                                                                             inscripcionencuesta=dato)
                        if respuesta.values("id").exists():
                            respuesta = respuesta.first()
                            ws.write(row_num, c, respuesta.respuesta, font_style2)
                            if not x.esta_vacia():
                                c += 1
                                ws.write(row_num, c, respuesta.respuestaporno, font_style2)

                        else:
                            respuesta = None
                            ws.write(row_num, c, '', font_style2)
                            if not x.esta_vacia():
                                c += 1
                                ws.write(row_num, c, '', font_style2)
                        c += 1
                    elif x.tipo == 2:
                        respuesta = RespuestaRangoEncuestaGrupoEstudiantes.objects.filter(status=True, pregunta=x,
                                                                                          inscripcionencuesta=dato)
                        if respuesta.values("id").exists():
                            respuesta = respuesta.first()
                            ws.write(row_num, c, respuesta.opcionrango.descripcion, font_style2)
                        else:
                            ws.write(row_num, c, '', font_style2)
                        c += 1
                    elif x.tipo in [3, 4]:
                        respuesta = RespuestaPreguntaEncuestaGrupoEstudiantes.objects.filter(status=True, pregunta=x,
                                                                                             inscripcionencuesta=dato)
                        if respuesta.values("id").exists():
                            respuesta = respuesta.first()
                            ws.write(row_num, c, respuesta.respuesta, font_style2)
                        else:
                            ws.write(row_num, c, '', font_style2)
                        c += 1
                    elif x.tipo == 5:
                        respuesta = dato.respuestacuadriculaencuestagrupoestudiantes_set.filter(status=True,
                                                                                                pregunta=x).first() if dato.respuestacuadriculaencuestagrupoestudiantes_set.filter(
                            status=True, pregunta=x).exists() else None
                        if respuesta is not None:
                            try:
                                int(respuesta.respuesta)
                                if OpcionCuadriculaEncuestaGrupoEstudiantes.objects.filter(status=True, pregunta=x,
                                                                                           id=respuesta.opcioncuadricula.id,
                                                                                           tipoopcion=2).first() == None:
                                    resp = 'Sin contestar'
                                else:
                                    resp = OpcionCuadriculaEncuestaGrupoEstudiantes.objects.filter(status=True,
                                                                                                   pregunta=x,
                                                                                                   id=respuesta.opcioncuadricula.id,
                                                                                                   tipoopcion=2).first().descripcion
                            except ValueError:
                                resp = respuesta.respuesta

                            ws.write(row_num, c, resp, font_style2)
                        else:
                            ws.write(row_num, c, '', font_style2)
                        c += 1
                    elif x.tipo == 6:
                        respuesta = dato.respuestamultipleencuestagrupoestudiantes_set.filter(status=True,
                                                                                              pregunta=x) if dato.respuestamultipleencuestagrupoestudiantes_set.values(
                            'id').filter(status=True, pregunta=x).exists() else None
                        if respuesta is not None:
                            w = 0
                            for rmult in respuesta:
                                ws.write(row_num + w, c, rmult.opcionmultiple.descripcion, font_style2)
                                # row_num += 1
                                w += 1
                            if limit < w and w > 0:
                                limit = w - 1
                        else:
                            ws.write(row_num, c, '', font_style2)
                        c += 1
            row_num += 1
            print('%s' % (row_num))

        except Exception as ex:
            print('error: %s' % (ex))
            pass

    wb.save(filename)
    print("FIN: ", filename)

    if eNotificacion:
        eNotificacion.en_proceso = False
        eNotificacion.titulo = f'Resultados de encuesta ID: {id}'
        eNotificacion.cuerpo = f'Finalizo generación de reporte de excel'
        eNotificacion.url = "{}reportes/encuestas/{}".format(MEDIA_URL, nombre)
        eNotificacion.save()
        send_user_notification(user=usernotify, payload={"head": eNotificacion.titulo,
                                                         "body": eNotificacion.cuerpo,
                                                         "action": "notificacion",
                                                         "timestamp": time.mktime(datetime.now().timetuple()),
                                                         "url": eNotificacion.url,
                                                         "btn_notificaciones": traerNotificaciones(None, None,
                                                                                                   ePersona),
                                                         "mensaje": f'Los resultados de la encuesta {id} han sido generados con exito'},
                               ttl=500)


def generate_precandidatos_beca(periodoactual, periodoanterior):
    error = False
    mensaje_ex = None
    with transaction.atomic():
        try:
            becatipos = BecaTipo.objects.filter(status=True, vigente=True)
            ID_GRUPO_VULNERABLE = 18
            EXCLUDES = []
            ePreInscripcionBecas = PreInscripcionBeca.objects.filter(periodo=periodoactual)
            EXCLUDES.extend(list(ePreInscripcionBecas.values_list("inscripcion__id", flat=True)))
            grupo = []
            for grupo_id in [4, 5]:
                grupo.extend(
                    lista_gruposocioeconomico_beca(periodoactual=periodoactual, periodoanterior=periodoanterior,
                                                   tipogrupo_id=grupo_id, excludes=EXCLUDES, limit=5))
            # inscripciones_grupo = Inscripcion.objects.filter(pk__in=grupo).exclude(persona_id__in=ePreInscripcionBecas.values("inscripcion__persona_id"))
            # actualizar_grupo = PreInscripcionBeca.objects.filter(becatipo=becatipos.filter(pk=ID_GRUPO_VULNERABLE).first(), periodo=periodoactual).exclude(inscripcion__in=inscripciones_grupo)
            # if DEBUG:
            #     actualizar_grupo.delete()
            inscripciones_grupo = grupo
            total = len(inscripciones_grupo)
            print(f"Total de inscripciones a registar {total}")
            contador = 0
            contadorNew = 0
            for inscripcion in inscripciones_grupo:
                contador += 1
                if not PreInscripcionBeca.objects.filter(inscripcion=inscripcion, periodo=periodoactual).exists():
                    preinscripcion = PreInscripcionBeca(inscripcion=inscripcion,
                                                        promedio=inscripcion.promediototal,
                                                        becatipo=becatipos.filter(pk=ID_GRUPO_VULNERABLE).first(),
                                                        periodo=periodoactual,
                                                        fecha=datetime.now().date())
                    preinscripcion.save()
                    preinscripcion.generar_requistosbecas()
                    contadorNew += 1
                    print(f"({total}/{contador}) Inscripcion: {inscripcion.__str__()} ->>> SE CREO")
                else:
                    print(f"({total}/{contador}) Inscripcion: {inscripcion.__str__()} ->>> NO SE CREO --YA EXISTE--")
            error = False
            print(f"Finaliza total que se crearon {contadorNew}")
        except Exception as ex:
            transaction.set_rollback(True)
            print(ex)
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
            error = True
            mensaje_ex = ex.__str__()


def arreglar_cagada_tipomatricula(periodo):
    eMatriculas = Matricula.objects.filter(nivel__periodo=periodo)
    total = len(eMatriculas.values("id"))
    print(f"{total}")
    contador = 0
    for eMatricula in eMatriculas:
        contador += 1
        itinerario = 0
        if not eMatricula.inscripcion.itinerario is None and eMatricula.inscripcion.itinerario > 0:
            itinerario = eMatricula.inscripcion.itinerario
        cantidad_seleccionadas = 0
        cursor = connections['default'].cursor()
        sql = "select am.nivelmalla_id, count(am.nivelmalla_id) as cantidad_materias_seleccionadas from sga_materiaasignada ma, sga_materia m, sga_asignaturamalla am where ma.status=true and ma.matricula_id=" + str(
            eMatricula.id) + " and m.status=true and m.id=ma.materia_id and am.status=true and am.id=m.asignaturamalla_id GROUP by am.nivelmalla_id, am.malla_id order by count(am.nivelmalla_id) desc, am.nivelmalla_id desc limit 1;"
        if itinerario > 0:
            sql = "select am.nivelmalla_id, count(am.nivelmalla_id) as cantidad_materias_seleccionadas from sga_materiaasignada ma, sga_materia m, sga_asignaturamalla am where ma.status=true and ma.matricula_id=" + str(
                eMatricula.id) + " and m.status=true and m.id=ma.materia_id and am.status=true and am.id=m.asignaturamalla_id and (am.itinerario=0 or am.itinerario=" + str(
                itinerario) + ") GROUP by am.nivelmalla_id, am.malla_id order by count(am.nivelmalla_id) desc, am.nivelmalla_id desc limit 1;"
        cursor.execute(sql)
        results = cursor.fetchall()
        nivel = 0
        for per in results:
            nivel = per[0]
            cantidad_seleccionadas = per[1]
        cantidad_nivel = 0
        eAsignaturaMallas = AsignaturaMalla.objects.filter(nivelmalla__id=nivel, status=True,
                                                           malla=eMatricula.inscripcion.mi_malla())
        if itinerario > 0:
            eAsignaturaMallas = eAsignaturaMallas.filter(Q(itinerario=0) | Q(itinerario=itinerario))

        for eAsignaturaMalla in eAsignaturaMallas:
            if Materia.objects.filter(nivel__periodo=eMatricula.nivel.periodo,
                                      asignaturamalla=eAsignaturaMalla).exists():
                if eMatricula.inscripcion.estado_asignatura(eAsignaturaMalla.asignatura) != 1:
                    cantidad_nivel += 1

        porcentaje_seleccionadas = int(round(
            Decimal((float(cantidad_nivel) * float(PORCIENTO_PERDIDA_PARCIAL_GRATUIDAD)) / 100).quantize(
                Decimal('.00')), 0))
        if (cantidad_seleccionadas < porcentaje_seleccionadas):
            tipo_matricula_ri = 2
        else:
            tipo_matricula_ri = 1
        eMatricula.grupo_socio_economico(tipo_matricula_ri)
        print(f"({contador}/{total}) --> {eMatricula} (actualizado)")


def crear_horario_clase_mooc():
    ePersonaRN = Profesor.objects.get(persona_id=83652)
    ePersonaLC = Profesor.objects.get(persona_id=83654)
    ePersonaIN = Profesor.objects.get(persona_id=83655)
    eTipoProfesorTutor = TipoProfesor.objects.get(pk=14)
    eTipoProfesorTeoria = TipoProfesor.objects.get(pk=1)
    dia = 6
    # horainicio_rn = datetime(2022, 1, 1, 15, 0, 0)
    # horafin_rn = datetime(2022, 1, 1, 15, 59, 59)
    # horainicio_in = datetime(2022, 1, 1, 16, 0, 0)
    # horafin_in = datetime(2022, 1, 1, 16, 59, 59)
    # horainicio_lc = datetime(2022, 1, 1, 17, 0, 0)
    # horafin_lc = datetime(2022, 1, 1, 17, 59, 59)
    eTurno_1 = Turno.objects.get(pk=552)
    eTurno_2 = Turno.objects.get(pk=553)
    eTurno_3 = Turno.objects.get(pk=554)
    ids_lc_en_linea = [60391, 60438, 60455, 60467, 60479, 60537, 58675, 58680, 58652, 58752, 60517, 60519, 60521, 60522,
                       56296, 56316, 56274, 59817, 59823, 59829, 56907, 57608, 57622, 57638, 57647, 57518, 57639, 58562,
                       58834, 58835, 58837, 58828, 58713, 58775, 60371, 57107, 57114, 60849, 60852, 60853, 60855, 60856,
                       56947]
    ids_rn_en_linea = [57463, 57508, 57509, 56968, 57048, 57510, 56944, 60370, 60372, 57002, 57600, 56268, 56269, 59250,
                       59256, 58640, 58821, 58824, 58825, 58694, 58684, 58626, 58739, 59526, 59531, 59536, 59541, 57352,
                       57427, 57582, 57399]
    ids_in_en_linea = [57216, 57250, 57387, 57511, 57687, 57513, 57587, 57489, 56866, 57588, 57226, 59134, 59140, 56917,
                       56919, 56925, 58831, 58832, 58838, 58840, 60433, 60451, 60463, 60475, 60487, 60499, 60505, 57452,
                       57551, 57689, 57481, 57470, 57115, 57117, 57122, 57134, 57142, 57144, 57506, 58879, 58538, 58630,
                       58692, 58613, 58808, 58606]
    ids_lc_presencial = [52664, 59789, 60143, 59790, 60183, 59802, 60253, 60130, 56120, 56213, 60124, 60282, 60286,
                         56687, 56735, 56808, 56720, 56732, 57850, 59850, 58023, 59840, 60319, 60322, 59857, 58025,
                         58962, 60298, 60303, 60308, 60313, 60810, 60813, 60544, 58090, 60534, 58465, 58404, 59770,
                         58300, 58051, 60770, 60749, 52577, 52790, 59862, 59867, 56441, 56497, 56695, 56570, 58299,
                         58319, 58172, 58080, 56414]
    ids_rn_presencial = [57808, 57902, 56595, 57806, 57920, 56430, 56706, 59240, 56704, 59193, 59198, 59204, 59209,
                         59213, 59224, 59235, 58289, 58282, 59983, 58146, 60571, 60613, 60774, 60753, 52559, 52777,
                         52638, 58401, 58409, 58439, 60878, 59917]
    ids_in_presencial = [52597, 52782, 52656, 60563, 60605, 60764, 60743, 56823, 56712, 60824, 56783, 60342, 60354,
                         57750, 57859, 58913, 59990, 60165, 60105, 56206, 56161, 60268, 59797, 59078, 59079, 59080,
                         56474, 56487, 56535, 59731, 59044, 59049, 59054, 60368, 58318, 58294, 60331, 60333, 60296,
                         60301, 60306, 60311, 60809, 60812, 58066, 58170, 58366, 60535, 60536, 59772, 60533, 60531,
                         56601, 59687, 56464, 56522]
    eMaterias_lc_en_linea = Materia.objects.filter(pk__in=ids_lc_en_linea)
    eMaterias_rn_en_linea = Materia.objects.filter(pk__in=ids_rn_en_linea)
    eMaterias_in_en_linea = Materia.objects.filter(pk__in=ids_in_en_linea)
    eMaterias_lc_presencial = Materia.objects.filter(pk__in=ids_lc_presencial)
    eMaterias_rn_presencial = Materia.objects.filter(pk__in=ids_rn_presencial)
    eMaterias_in_presencial = Materia.objects.filter(pk__in=ids_in_presencial)

    for eMateria in eMaterias_lc_en_linea:
        eMateria.tipomateria = 3
        eMateria.save()
        eProfesorMaterias = ProfesorMateria.objects.filter(materia=eMateria, tipoprofesor=eTipoProfesorTutor)
        if eProfesorMaterias.values("id").exists():
            eProfesorMateria = eProfesorMaterias[0]
            eProfesorMateria.profesor = ePersonaLC
            eProfesorMateria.desde = eMateria.inicio
            eProfesorMateria.hasta = eMateria.fin
            eProfesorMateria.principal = True
            eProfesorMateria.save()
        else:
            eProfesorMateria = ProfesorMateria(materia=eMateria,
                                               tipoprofesor=eTipoProfesorTutor,
                                               profesor=ePersonaLC,
                                               desde=eMateria.inicio,
                                               hasta=eMateria.fin,
                                               principal=True
                                               )
            eProfesorMateria.save()
        Clase.objects.filter(materia=eMateria).delete()
        # eClase = Clase(materia=eMateria,
        #                turno=eTurno_1,
        #                profesor=eProfesorMateria.profesor,
        #                tipohorario=10,
        #                aula_id=218,
        #                dia=dia,
        #                inicio=eMateria.inicio,
        #                fin=eMateria.fin,
        #                tipoprofesor=eProfesorMateria.tipoprofesor,
        #                subirenlace=False)
        # eClase.save()

    for eMateria in eMaterias_rn_en_linea:
        eMateria.tipomateria = 3
        eMateria.save()
        eProfesorMaterias = ProfesorMateria.objects.filter(materia=eMateria, tipoprofesor=eTipoProfesorTutor)
        if eProfesorMaterias.values("id").exists():
            eProfesorMateria = eProfesorMaterias[0]
            eProfesorMateria.profesor = ePersonaRN
            eProfesorMateria.desde = eMateria.inicio
            eProfesorMateria.hasta = eMateria.fin
            eProfesorMateria.principal = True
            eProfesorMateria.save()
        else:
            eProfesorMateria = ProfesorMateria(materia=eMateria,
                                               tipoprofesor=eTipoProfesorTutor,
                                               profesor=ePersonaRN,
                                               desde=eMateria.inicio,
                                               hasta=eMateria.fin,
                                               principal=True
                                               )
            eProfesorMateria.save()
        Clase.objects.filter(materia=eMateria).delete()
        # eClase = Clase(materia=eMateria,
        #                turno=eTurno_2,
        #                profesor=eProfesorMateria.profesor,
        #                tipohorario=10,
        #                aula_id=218,
        #                dia=dia,
        #                inicio=eMateria.inicio,
        #                fin=eMateria.fin,
        #                tipoprofesor=eProfesorMateria.tipoprofesor,
        #                subirenlace=False)
        # eClase.save()

    for eMateria in eMaterias_in_en_linea:
        eMateria.tipomateria = 3
        eMateria.save()
        eProfesorMaterias = ProfesorMateria.objects.filter(materia=eMateria, tipoprofesor=eTipoProfesorTutor)
        if eProfesorMaterias.values("id").exists():
            eProfesorMateria = eProfesorMaterias[0]
            eProfesorMateria.profesor = ePersonaIN
            eProfesorMateria.desde = eMateria.inicio
            eProfesorMateria.hasta = eMateria.fin
            eProfesorMateria.principal = True
            eProfesorMateria.save()
        else:
            eProfesorMateria = ProfesorMateria(materia=eMateria,
                                               tipoprofesor=eTipoProfesorTutor,
                                               profesor=ePersonaIN,
                                               desde=eMateria.inicio,
                                               hasta=eMateria.fin,
                                               principal=True
                                               )
            eProfesorMateria.save()
        Clase.objects.filter(materia=eMateria).delete()
        # eClase = Clase(materia=eMateria,
        #                turno=eTurno_3,
        #                profesor=eProfesorMateria.profesor,
        #                tipohorario=10,
        #                aula_id=218,
        #                dia=dia,
        #                inicio=eMateria.inicio,
        #                fin=eMateria.fin,
        #                tipoprofesor=eProfesorMateria.tipoprofesor,
        #                subirenlace=False)
        # eClase.save()

    for eMateria in eMaterias_lc_presencial:
        eMateria.tipomateria = 3
        eMateria.save()
        eProfesorMaterias = ProfesorMateria.objects.filter(materia=eMateria, tipoprofesor=eTipoProfesorTeoria)
        if eProfesorMaterias.values("id").exists():
            eProfesorMateria = eProfesorMaterias[0]
            eProfesorMateria.profesor = ePersonaLC
            eProfesorMateria.desde = eMateria.inicio
            eProfesorMateria.hasta = eMateria.fin
            eProfesorMateria.principal = True
            eProfesorMateria.save()
        else:
            eProfesorMateria = ProfesorMateria(materia=eMateria,
                                               tipoprofesor=eTipoProfesorTeoria,
                                               profesor=ePersonaLC,
                                               desde=eMateria.inicio,
                                               hasta=eMateria.fin,
                                               principal=True
                                               )
            eProfesorMateria.save()
        Clase.objects.filter(materia=eMateria).delete()
        # eClase = Clase(materia=eMateria,
        #                turno=eTurno_1,
        #                profesor=eProfesorMateria.profesor,
        #                tipohorario=10,
        #                aula_id=218,
        #                dia=dia,
        #                inicio=eMateria.inicio,
        #                fin=eMateria.fin,
        #                tipoprofesor=eProfesorMateria.tipoprofesor,
        #                subirenlace=False)
        # eClase.save()

    for eMateria in eMaterias_rn_presencial:
        eMateria.tipomateria = 3
        eMateria.save()
        eProfesorMaterias = ProfesorMateria.objects.filter(materia=eMateria, tipoprofesor=eTipoProfesorTeoria)
        if eProfesorMaterias.values("id").exists():
            eProfesorMateria = eProfesorMaterias[0]
            eProfesorMateria.profesor = ePersonaRN
            eProfesorMateria.desde = eMateria.inicio
            eProfesorMateria.hasta = eMateria.fin
            eProfesorMateria.principal = True
            eProfesorMateria.save()
        else:
            eProfesorMateria = ProfesorMateria(materia=eMateria,
                                               tipoprofesor=eTipoProfesorTeoria,
                                               profesor=ePersonaRN,
                                               desde=eMateria.inicio,
                                               hasta=eMateria.fin,
                                               principal=True
                                               )
            eProfesorMateria.save()
        Clase.objects.filter(materia=eMateria).delete()
        # eClase = Clase(materia=eMateria,
        #                turno=eTurno_2,
        #                profesor=eProfesorMateria.profesor,
        #                tipohorario=10,
        #                aula_id=218,
        #                dia=dia,
        #                inicio=eMateria.inicio,
        #                fin=eMateria.fin,
        #                tipoprofesor=eProfesorMateria.tipoprofesor,
        #                subirenlace=False)
        # eClase.save()

    for eMateria in eMaterias_in_presencial:
        eMateria.tipomateria = 3
        eMateria.save()
        eProfesorMaterias = ProfesorMateria.objects.filter(materia=eMateria, tipoprofesor=eTipoProfesorTeoria)
        if eProfesorMaterias.values("id").exists():
            eProfesorMateria = eProfesorMaterias[0]
            eProfesorMateria.profesor = ePersonaIN
            eProfesorMateria.desde = eMateria.inicio
            eProfesorMateria.hasta = eMateria.fin
            eProfesorMateria.principal = True
            eProfesorMateria.save()
        else:
            eProfesorMateria = ProfesorMateria(materia=eMateria,
                                               tipoprofesor=eTipoProfesorTeoria,
                                               profesor=ePersonaIN,
                                               desde=eMateria.inicio,
                                               hasta=eMateria.fin,
                                               principal=True
                                               )
            eProfesorMateria.save()
        Clase.objects.filter(materia=eMateria).delete()
        # eClase = Clase(materia=eMateria,
        #                turno=eTurno_3,
        #                profesor=eProfesorMateria.profesor,
        #                tipohorario=10,
        #                aula_id=218,
        #                dia=dia,
        #                inicio=eMateria.inicio,
        #                fin=eMateria.fin,
        #                tipoprofesor=eProfesorMateria.tipoprofesor,
        #                subirenlace=False)
        # eClase.save()


def crear_requisitos_ingreso_aic(periodo_id=153):
    eMaterias = Materia.objects.filter(nivel__periodo_id=periodo_id, asignaturamalla__validarequisitograduacion=True)
    print(f"Materias # {len(eMaterias.values('id'))}")
    for eMateria in eMaterias:
        print(f"Inicia -- Materia: {eMateria.__str__()}")
        eMateria.crear_actualizar_requisitos_uic(actualizar=True)
        print(f"Fin -- Materia: {eMateria.__str__()}")


# def ActualizarTestRestrinccionMoodle(tareaid, persona):
#     from sga.models import TestSilaboSemanal
#     from django.db import connections
#
#     eTestSilaboSemanal = TestSilaboSemanal.objects.get(pk=tareaid)
#     eMateria = eTestSilaboSemanal.silabosemanal.silabo.materia
#     eMateriaAsignadas = eMateria.materiaasignada_set.select_related().filter(status=True, retiramateria=False)
#
#
#     if eMateria.coordinacion():
#         if eMateria.coordinacion().id == 9:
#             cursor_verbose = 'db_moodle_virtual'
#         else:
#             cursor_verbose = 'moodle_db'
#     else:
#         cursor_verbose = 'moodle_db'
#     if eMateria.idcursomoodle == 0:
#         return False, u"Materia no tiene creado el curso en Moodle"
#     availability = '{"op":"&","c":[{"op":"&","c":[{"type":"profile","sf":"email","op":"isequalto","v":"crodriguezn@unemi.edu.ec"},{"type":"date","d":">=","t":1659985200},{"type":"date","d":"<","t":1659992400}]}],"showc":[true]}'
#
#
#
#     with transaction.atomic(using=cursor_verbose):
#         try:
#             cursoid = eMateria.idcursomoodle
#             cursor = None
#             conexion = None
#             if eMateria.coordinacion():
#                 if eMateria.coordinacion().id == 9:
#                     # cursor = connections['db_moodle_virtual'].cursor()
#                     conexion = connections['db_moodle_virtual']
#                 else:
#                     # cursor = connections['moodle_db'].cursor()
#                     conexion = connections['moodle_db']
#             else:
#                 # cursor = connections['moodle_db'].cursor()
#                 conexion = connections['moodle_db']
#
#             cursor = conexion.cursor()
#             sql = """select id from mooc_course_sections WHERE course=%s AND SECTION='%s' """ % (cursoid, 4)
#             cursor.execute(sql)
#             buscar = cursor.fetchall()
#             if not buscar:
#                 return False, u"La configuración de secciones de moodle es diferente a la establecida"
#             section = buscar[0][0]
#             intro = ""
#             horadesde = 0
#             minutodesde = 0
#             horahasta = 23
#             minutohasta = 59
#             if eTestSilaboSemanal.horadesde:
#                 horadesde = eTestSilaboSemanal.horadesde.hour
#                 minutodesde = eTestSilaboSemanal.horadesde.minute
#
#             if eTestSilaboSemanal.horahasta:
#                 horahasta = eTestSilaboSemanal.horahasta.hour
#                 minutohasta = eTestSilaboSemanal.horahasta.minute
#
#             fecha = int(time.mktime(datetime.now().timetuple()))
#             fechadesde = datetime(eTestSilaboSemanal.fechadesde.date().year, eTestSilaboSemanal.fechadesde.date().month,
#                                   eTestSilaboSemanal.fechadesde.date().day, horadesde, minutodesde)
#             fechadesde = int(time.mktime(fechadesde.timetuple()))
#             fechahasta = datetime(eTestSilaboSemanal.fechahasta.date().year, eTestSilaboSemanal.fechahasta.date().month,
#                                   eTestSilaboSemanal.fechahasta.date().day, horahasta, minutohasta)
#             fechahasta = int(time.mktime(fechahasta.timetuple()))
#             limitetiempo = eTestSilaboSemanal.tiempoduracion * 60
#             notamaxima = 0
#             if eTestSilaboSemanal.calificar:
#                 notamaxima = eTestSilaboSemanal.detallemodelo.notamaxima
#
#             if eTestSilaboSemanal.navegacion == 1:
#                 navegacion = 'free'
#             else:
#                 navegacion = 'sequential'
#
#             if eTestSilaboSemanal.idtestmoodle == 0:
#                 return False, u"La configuración de secciones de moodle es diferente a la establecida"
#             else:
#                 instanceid = eTestSilaboSemanal.idtestmoodle
#
#                 sql = """select id from mooc_course_modules WHERE course=%s AND module=17 and instance='%s' """ % (cursoid, instanceid)
#                 cursor.execute(sql)
#                 buscar = cursor.fetchall()
#                 course_modules = buscar[0][0]
#
#                 sql = """update mooc_course_modules set section='%s' where course=%s and module=17 and instance=%s""" % (section, cursoid, instanceid)
#                 cursor.execute(sql)
#
#                 sql = u"UPDATE mooc_course SET cacherev = %s WHERE id = %s" % (fecha, cursoid)
#                 cursor.execute(sql)
#
#                 sql = """select id from mooc_context WHERE contextlevel=70 AND instanceid='%s' """ % (course_modules)
#                 cursor.execute(sql)
#                 buscar = cursor.fetchall()
#                 if not buscar:
#                     sql = """INSERT INTO mooc_context (contextlevel,instanceid,depth,path,locked) VALUES('70','%s','0',NULL,'0')""" % (course_modules)
#                     cursor.execute(sql)
#
#                 # sql = """select array_to_string(array_agg(id),',') from mooc_course_modules where course=%s and module=17 and instance in (select id from mooc_quiz where course=%s)""" % (cursoid, cursoid)
#                 sql = """select array_to_string(array_agg(id),',') from mooc_course_modules where deletioninprogress=0 and course=%s and section=%s""" % (cursoid, section)
#                 cursor.execute(sql)
#                 buscar = cursor.fetchall()
#                 course_modules = buscar[0][0]
#
#                 sql = """UPDATE mooc_course_sections SET sequence = '%s' WHERE course = %s and section=4""" % (course_modules, cursoid)
#                 cursor.execute(sql)
#
#                 sql = """update mooc_course_modules set section='%s' where course=%s and module=17 and instance=%s""" % (section, cursoid, instanceid)
#                 cursor.execute(sql)
#
#                 sql = u"UPDATE mooc_course SET cacherev = %s WHERE id = %s" % (fecha, cursoid)
#                 cursor.execute(sql)
#
#             if tarea.idtestmoodle > 0:
#                 tarea.estado_id = 4
#                 tarea.save()
#             return True, u"Recurso migrado a Moodle"
#         except Exception as ex:
#             transaction.set_rollback(True, using=cursor_verbose)
#             return False, "%s - %s" % (ex.__str__(), sys.exc_info()[-1].tb_lineno)
#         finally:
#             cursor.close()


# descarga_resultado_encuesta(24, 'reporte_encuesta_24')
# descarga_resultado_encuesta(25, 'reporte_encuesta_25')
# descarga_resultado_encuesta(32, 'reporte_encuesta_32')


# periodoactual = Periodo.objects.get(status=True, id=126)
# periodoanterior = Periodo.objects.get(status=True, id=119)
# arreglar_cagada_tipomatricula(periodoactual)


# eMatriculas = Matricula.objects.filter(nivel__periodo_id=153, nivel_id=740)
# print(f"Total: {len(eMatriculas.values('id'))}")
# eMatriculas.delete()


def ajuste_recalcular_rubros(idp):
    # eMatriculas = Matricula.objects.filter(status=True, nivel__periodo_id=idp, rubro__status=True, nivel_id__in=[684, 687], fecha__gte=datetime(2022, 10, 19, 0, 0, 0)).distinct()
    eMatriculas = Matricula.objects.filter(status=True, nivel__periodo_id=idp, rubro__status=True).distinct()
    # eMatriculas = eMatriculas.filter(pk=515160)
    total = len(eMatriculas.values("id"))
    print(f"Total encontrados: {total}")
    contador = 0
    errors = []
    for eMatricula in eMatriculas:
        exError = False
        with transaction.atomic():
            try:
                eMatricula.inscripcion.actualiza_estado_matricula()
                valid, msg, aData = get_tipo_matricula(None, eMatricula)
                if not valid:
                    raise NameError(msg)
                cantidad_nivel = aData['cantidad_nivel']
                porcentaje_perdidad_parcial_gratuidad = aData['porcentaje_perdidad_parcial_gratuidad']
                cantidad_seleccionadas = aData['cantidad_seleccionadas']
                porcentaje_seleccionadas = int(round(
                    Decimal((float(cantidad_nivel) * float(porcentaje_perdidad_parcial_gratuidad)) / 100).quantize(
                        Decimal('.00')), 0))
                if (cantidad_seleccionadas < porcentaje_seleccionadas):
                    eMatricula.grupo_socio_economico(2)
                else:
                    eMatricula.grupo_socio_economico(1)
                if not eMatricula.tiene_pagos_matricula():
                    aranceldiferido = eMatricula.aranceldiferido
                    eMatricula.agregacion_aux(request=None)
                    if aranceldiferido == 1:
                        ePeriodoMatricula = None
                        if Rubro.objects.filter(matricula=eMatricula, tipo_id=RUBRO_ARANCEL, status=True).count() == 1:
                            print(eMatricula)
                            if PeriodoMatricula.objects.values('id').filter(status=True, activo=True,
                                                                            periodo=eMatricula.nivel.periodo).exists():
                                ePeriodoMatricula = PeriodoMatricula.objects.get(status=True, activo=True,
                                                                                 periodo=eMatricula.nivel.periodo)
                            if Rubro.objects.values('id').filter(matricula=eMatricula, tipo_id=RUBRO_ARANCEL,
                                                                 status=True).exists():
                                arancel = Rubro.objects.filter(matricula=eMatricula, tipo_id=RUBRO_ARANCEL,
                                                               status=True).first()
                                nombrearancel = arancel.nombre
                                valorarancel = Decimal(arancel.valortotal).quantize(Decimal('.01'))
                                num_cuotas = 3
                                try:
                                    valor_cuota_mensual = (valorarancel / num_cuotas).quantize(Decimal('.01'))
                                except ZeroDivisionError:
                                    valor_cuota_mensual = 0
                                if valor_cuota_mensual == 0:
                                    raise NameError(u"No se puede procesar el registro.")
                                eRubroMatricula = Rubro.objects.filter(matricula=eMatricula, tipo_id=RUBRO_MATRICULA)[0]
                                eRubroMatricula.relacionados = None
                                eRubroMatricula.save()
                                lista = []
                                c = 0
                                for r in ePeriodoMatricula.fecha_cuotas_rubro().values('fecha').distinct():
                                    c += 1
                                    lista.append([c, valor_cuota_mensual, r['fecha']])
                                for item in lista:
                                    rubro = Rubro(tipo_id=RUBRO_ARANCEL,
                                                  persona=eMatricula.inscripcion.persona,
                                                  relacionados=eRubroMatricula,
                                                  matricula=eMatricula,
                                                  nombre=nombrearancel,
                                                  cuota=item[0],
                                                  fecha=datetime.now().date(),
                                                  fechavence=item[2],
                                                  valor=item[1],
                                                  iva_id=1,
                                                  valoriva=0,
                                                  valortotal=item[1],
                                                  saldo=item[1],
                                                  cancelado=False)
                                    rubro.save()
                                arancel.delete()
                                if ePeriodoMatricula.valida_rubro_acta_compromiso:
                                    isResult, message = eMatricula.generar_actacompromiso_matricula_pregrado()
                                    if not isResult:
                                        raise NameError(message)
                                    eMatricula.aranceldiferido = 1
                                    eMatricula.actacompromiso = message
                                    eMatricula.save()
                    # eMatricula.calcula_nivel()
                    # itinerario = 0
                    # if not eMatricula.inscripcion.itinerario is None and eMatricula.inscripcion.itinerario > 0:
                    #     itinerario = eMatricula.inscripcion.itinerario
                    # cantidad_seleccionadas = 0
                    # cursor = connections['default'].cursor()
                    # sql = "select am.nivelmalla_id, count(am.nivelmalla_id) as cantidad_materias_seleccionadas from sga_materiaasignada ma, sga_materia m, sga_asignaturamalla am where ma.status=true and ma.matricula_id=" + str(eMatricula.id) + " and m.status=true and m.id=ma.materia_id and am.status=true and am.id=m.asignaturamalla_id GROUP by am.nivelmalla_id, am.malla_id order by count(am.nivelmalla_id) desc, am.nivelmalla_id desc limit 1;"
                    # if itinerario > 0:
                    #     sql = "select am.nivelmalla_id, count(am.nivelmalla_id) as cantidad_materias_seleccionadas from sga_materiaasignada ma, sga_materia m, sga_asignaturamalla am where ma.status=true and ma.matricula_id=" + str(eMatricula.id) + " and m.status=true and m.id=ma.materia_id and am.status=true and am.id=m.asignaturamalla_id and (am.itinerario=0 or am.itinerario=" + str(itinerario) + ") GROUP by am.nivelmalla_id, am.malla_id order by count(am.nivelmalla_id) desc, am.nivelmalla_id desc limit 1;"
                    # cursor.execute(sql)
                    # results = cursor.fetchall()
                    # nivel = 0
                    # for per in results:
                    #     nivel = per[0]
                    #     cantidad_seleccionadas = per[1]
                    # cantidad_nivel = 0
                    # eAsignaturaMallas = AsignaturaMalla.objects.filter(nivelmalla__id=nivel, status=True, malla=eMatricula.inscripcion.mi_malla())
                    # if itinerario > 0:
                    #     eAsignaturaMallas = eAsignaturaMallas.filter(Q(itinerario=0) | Q(itinerario=itinerario))
                    #
                    # for eAsignaturaMalla in eAsignaturaMallas:
                    #     if Materia.objects.filter(nivel__periodo=eMatricula.nivel.periodo, asignaturamalla=eAsignaturaMalla).exists():
                    #         if eMatricula.inscripcion.estado_asignatura(eAsignaturaMalla.asignatura) != 1:
                    #             cantidad_nivel += 1
                    #
                    # porcentaje_seleccionadas = int(round(Decimal((float(cantidad_nivel) * float(PORCIENTO_PERDIDA_PARCIAL_GRATUIDAD)) / 100).quantize(Decimal('.00')), 0))
                    # if (cantidad_seleccionadas < porcentaje_seleccionadas):
                    #     tipo_matricula_ri = 2
                    # else:
                    #     tipo_matricula_ri = 1
                    # eMatricula.grupo_socio_economico(tipo_matricula_ri)
                    # eMatriculaGrupoSocioEconomicos = eMatricula.matriculagruposocioeconomico_set.all()
                    # if eMatriculaGrupoSocioEconomicos.values('id').exists():
                    #     eMatriculaGrupoSocioEconomico = eMatriculaGrupoSocioEconomicos[0]
                    #     eMatriculaGrupoSocioEconomico.estado_gratuidad = eMatricula.inscripcion.estado_gratuidad
                    #     eMatriculaGrupoSocioEconomico.save()
                eMatricula.save()
            except Exception as ex:
                exError = True
                transaction.set_rollback(True)
                errors.append(eMatricula.pk)
                print(f"Ocurrio un error en la matricula {eMatricula.__str__()}")
        # if not exError:
        #     with transaction.atomic():
        #         try:
        #             eMatricula.inscripcion.actualiza_estado_matricula()
        #             eMatricula.save()
        #
        #             # eMatricula.calcula_nivel()
        #             if not eMatricula.tiene_pagos_matricula():
        #                 aranceldiferido = eMatricula.aranceldiferido
        #                 contador += 1
        #                 print(f"Inicia --> ({total}/{contador}) - {eMatricula.__str__()}")
        #                 # eMatricula.calcula_nivel()
        #                 eMatricula.agregacion_aux(None)
        #                 print(f"Finalizo --> ({total}/{contador}) - {eMatricula.__str__()}")
        #                 if aranceldiferido == 1:
        #                     ePeriodoMatricula = None
        #                     if PeriodoMatricula.objects.values('id').filter(status=True, activo=True, periodo=eMatricula.nivel.periodo).exists():
        #                         ePeriodoMatricula = PeriodoMatricula.objects.get(status=True, activo=True, periodo=eMatricula.nivel.periodo)
        #                     if Rubro.objects.values('id').filter(matricula=eMatricula, tipo_id=RUBRO_ARANCEL, status=True).exists():
        #                         arancel = Rubro.objects.filter(matricula=eMatricula, tipo_id=RUBRO_ARANCEL, status=True).first()
        #                         nombrearancel = arancel.nombre
        #                         valorarancel = Decimal(arancel.valortotal).quantize(Decimal('.01'))
        #                         num_cuotas = 3
        #                         try:
        #                             valor_cuota_mensual = (valorarancel / num_cuotas).quantize(Decimal('.01'))
        #                         except ZeroDivisionError:
        #                             valor_cuota_mensual = 0
        #                         if valor_cuota_mensual == 0:
        #                             raise NameError(u"No se puede procesar el registro.")
        #                         eRubroMatricula = Rubro.objects.filter(matricula=eMatricula, tipo_id=RUBRO_MATRICULA)[0]
        #                         eRubroMatricula.relacionados = None
        #                         eRubroMatricula.save()
        #                         lista = []
        #                         c = 0
        #                         for r in ePeriodoMatricula.fecha_cuotas_rubro().values('fecha').distinct():
        #                             c += 1
        #                             lista.append([c, valor_cuota_mensual, r['fecha']])
        #                         for item in lista:
        #                             rubro = Rubro(tipo_id=RUBRO_ARANCEL,
        #                                           persona=eMatricula.inscripcion.persona,
        #                                           relacionados=eRubroMatricula,
        #                                           matricula=eMatricula,
        #                                           nombre=nombrearancel,
        #                                           cuota=item[0],
        #                                           fecha=datetime.now().date(),
        #                                           fechavence=item[2],
        #                                           valor=item[1],
        #                                           iva_id=1,
        #                                           valoriva=0,
        #                                           valortotal=item[1],
        #                                           saldo=item[1],
        #                                           cancelado=False)
        #                             rubro.save()
        #                         arancel.delete()
        #                         if ePeriodoMatricula.valida_rubro_acta_compromiso:
        #                             isResult, message = eMatricula.generar_actacompromiso_matricula_pregrado()
        #                             if not isResult:
        #                                 raise NameError(message)
        #                             eMatricula.aranceldiferido = 1
        #                             eMatricula.actacompromiso = message
        #                             eMatricula.save()
        #             valid, msg, aData = get_tipo_matricula(None, eMatricula)
        #             if not valid:
        #                 raise NameError(msg)
        #             cantidad_nivel = aData['cantidad_nivel']
        #             porcentaje_perdidad_parcial_gratuidad = aData['porcentaje_perdidad_parcial_gratuidad']
        #             cantidad_seleccionadas = aData['cantidad_seleccionadas']
        #             porcentaje_seleccionadas = int(round(Decimal((float(cantidad_nivel) * float(porcentaje_perdidad_parcial_gratuidad)) / 100).quantize(Decimal('.00')), 0))
        #             if (cantidad_seleccionadas < porcentaje_seleccionadas):
        #                 eMatricula.grupo_socio_economico(2)
        #             else:
        #                 eMatricula.grupo_socio_economico(1)
        #         except Exception as ex:
        #             transaction.set_rollback(True)
        #             errors.append(eMatricula.pk)
        #             print(f"Ocurrio un error en la matricula {eMatricula.__str__()}")
    print("---------------------------------------------------------------------------------------------------------")
    print(f"Error de id de matricula {errors}")


def ajuste_estado_matricula():
    periodoadmision = Periodo.objects.get(pk=158)
    periodopregrado = Periodo.objects.get(pk=153)
    config = My_ConfigMatriculacionPrimerNivel.objects.get(periodoadmision=periodoadmision,
                                                           periodopregrado=periodopregrado)
    for eMatriculacionPrimerNivelCarrera in My_MatriculacionPrimerNivelCarrera.objects.filter(configuracion=config):
        eCarreraAdmision = eMatriculacionPrimerNivelCarrera.carreraadmision
        eCarreraPregrado = eMatriculacionPrimerNivelCarrera.carrerapregrado
        print(f"Carrera Admision: {eCarreraAdmision.__str__()} -- Carrera Pregrado: {eCarreraPregrado.__str__()}")
        niveles_admision = Nivel.objects.filter(periodo=eMatriculacionPrimerNivelCarrera.configuracion.periodoadmision)
        alumnos_admision = My_MatriculaPregrado.objects.filter(status=True, aprobado=True,
                                                               nivel__in=niveles_admision.values_list('id'),
                                                               inscripcion__carrera=eCarreraAdmision)
        print(f"            ---------- Total encontrados: {len(alumnos_admision.values('id'))}")
        # if DEBUG:
        #     alumnos_admision = alumnos_admision.filter(pk=507510)
        for matriculaadmision in alumnos_admision:
            eInscripcionAdmision = matriculaadmision.inscripcion
            print(f"            ---------- Persona: {eInscripcionAdmision.persona.__str__()}")
            if Inscripcion.objects.values("id").filter(persona=eInscripcionAdmision.persona, carrera=eCarreraPregrado):
                eInscripcionPregrado = \
                Inscripcion.objects.filter(persona=eInscripcionAdmision.persona, carrera=eCarreraPregrado)[0]
                eInscripcionPregrado.estado_gratuidad = 1
                eInscripcionPregrado.porcentaje_perdida_gratuidad = 0
                if eInscripcionAdmision.estado_gratuidad == 3:
                    ePerdidaGratuidadAdmisions = PerdidaGratuidad.objects.filter(inscripcion=eInscripcionAdmision,
                                                                                 status=True)
                    ePerdidaGratuidadAdmision = None
                    observacion = None
                    if ePerdidaGratuidadAdmisions.values("id").exists():
                        observacion = 'Reportado por la SENESCYT'
                        ePerdidaGratuidadAdmision = ePerdidaGratuidadAdmisions[0]
                        if ePerdidaGratuidadAdmision.observacion:
                            observacion = ePerdidaGratuidadAdmision.observacion
                    if ePerdidaGratuidadAdmision is None:
                        if PerdidaGratuidad.objects.values('id').filter(inscripcion=eInscripcionPregrado,
                                                                        status=True).exists():
                            PerdidaGratuidad.objects.filter(inscripcion=eInscripcionPregrado, status=True).delete()
                        eInscripcionPregrado.estado_gratuidad = 1
                        eInscripcionPregrado.porcentaje_perdida_gratuidad = 0
                    else:
                        if not PerdidaGratuidad.objects.values('id').filter(inscripcion=eInscripcionPregrado,
                                                                            status=True).exists():
                            ePerdidaGratuidad = PerdidaGratuidad(inscripcion=eInscripcionPregrado,
                                                                 motivo=1,
                                                                 observacion=observacion)
                            ePerdidaGratuidad.save()
                            eInscripcionPregrado.estado_gratuidad = 3
                            eInscripcionPregrado.porcentaje_perdida_gratuidad = 100

                eInscripcionPregrado.save()
                eMatriculas = Matricula.objects.filter(inscripcion=eInscripcionPregrado, nivel__periodo=periodopregrado)
                if eMatriculas.values("id").exists():
                    eMatricula = eMatriculas[0]
                    eMatricula.agregacion_aux(None)


# for eCostoOptimoMalla in CostoOptimoMalla.objects.filter(periodo_id=153):
#     eCostoOptimoMalla.save()
#     eCostoOptimoMalla.crear_editar_calculo_niveles()
#     print(f"Finalizo ajuste de calculo de malla: {eCostoOptimoMalla.__str__()}")


def eliminar_registro_record_materia(idm):
    eMateria = Materia.objects.get(pk=idm)
    print(f"Materia: {eMateria.__str__()}")
    eRecordAcademicos = RecordAcademico.objects.filter(materiaregular=eMateria,
                                                       fecha_creacion__date=datetime(2022, 10, 19, 0, 0, 0).date())
    eInscripciones = Inscripcion.objects.filter(pk__in=eRecordAcademicos.values_list("inscripcion__id", flat=True))
    eHistoricoRecordAcademicos = HistoricoRecordAcademico.objects.filter(materiaregular=eMateria,
                                                                         fecha_creacion__date=datetime(2022, 10, 19, 0,
                                                                                                       0, 0).date())
    print(f"Total a eliminar historico de record académico: {len(eHistoricoRecordAcademicos.values('id'))}")
    eHistoricoRecordAcademicos.delete()
    print(f"Total a eliminar record académico: {len(eRecordAcademicos.values('id'))}")
    eRecordAcademicos.delete()
    print(f"-- EMPIZA A ACTUALIZAR ESTADOS DE INSCRIPCIONES")
    total = len(eInscripciones.values("id"))
    contador = 0
    for eInscripcion in eInscripciones:
        contador += 1
        print(f"({total}/{contador}) -> Inscripcion: {eInscripcion.__str__()}")
        eInscripcion.actualiza_estado_matricula()
    print(f"----- TERMINO PROCESO---------")


def doble_carrera(ePeriodo):
    erorres = []
    eMatriculas = Matricula.objects.filter(nivel__periodo=ePeriodo)
    ePersonas = Persona.objects.filter(pk__in=eMatriculas.values_list("inscripcion__persona__id", flat=True)).distinct()
    total_persona = len(ePersonas.values("id"))
    contador_persona = 0
    total_reajuste = 0
    for ePersona in ePersonas:
        contador_persona += 1
        print(f"({total_persona}/{contador_persona}) Persona --> {ePersona.__str__()}")
        if len(eMatriculas.values("id").filter(inscripcion__persona=ePersona)) > 1:
            eInscripciones = Inscripcion.objects.filter(
                pk__in=eMatriculas.values_list("inscripcion__id", flat=True).filter(
                    inscripcion__persona=ePersona)).order_by('-fechainiciocarrera', '-pk')
            total = len(eInscripciones.values("id"))
            contador = 0
            for eInscripcion in eInscripciones:
                contador += 1
                if contador < total:
                    with transaction.atomic():
                        try:
                            if not PerdidaGratuidad.objects.values("id").filter(inscripcion=eInscripcion,
                                                                                status=True).exists():
                                total_reajuste += 1
                                ePerdidaGratuidad = PerdidaGratuidad(inscripcion=eInscripcion,
                                                                     motivo=1,
                                                                     cupo_aceptado_senescyt=True,
                                                                     segunda_carrera_raes=True,
                                                                     observacion=u"Se encuentra cursando segunda carrera")
                                ePerdidaGratuidad.save()
                                eInscripcion.gratuidad = False
                                eInscripcion.estado_gratuidad = 3
                                eInscripcion.porcentaje_perdida_gratuidad = 100
                                eInscripcion.save()
                            eMatricula = eMatriculas.filter(inscripcion=eInscripcion)[0]
                            if not eMatricula.tiene_pagos_matricula():
                                aranceldiferido = eMatricula.aranceldiferido
                                eMatricula.agregacion_aux(request=None)
                                if aranceldiferido == 1:
                                    ePeriodoMatricula = None
                                    if PeriodoMatricula.objects.values('id').filter(status=True, activo=True,
                                                                                    periodo=eMatricula.nivel.periodo).exists():
                                        ePeriodoMatricula = PeriodoMatricula.objects.get(status=True, activo=True,
                                                                                         periodo=eMatricula.nivel.periodo)
                                    if Rubro.objects.values('id').filter(matricula=eMatricula, tipo_id=RUBRO_ARANCEL,
                                                                         status=True).exists():
                                        arancel = Rubro.objects.filter(matricula=eMatricula, tipo_id=RUBRO_ARANCEL,
                                                                       status=True).first()
                                        nombrearancel = arancel.nombre
                                        valorarancel = Decimal(arancel.valortotal).quantize(Decimal('.01'))
                                        num_cuotas = 3
                                        try:
                                            valor_cuota_mensual = (valorarancel / num_cuotas).quantize(Decimal('.01'))
                                        except ZeroDivisionError:
                                            valor_cuota_mensual = 0
                                        if valor_cuota_mensual == 0:
                                            raise NameError(u"No se puede procesar el registro.")
                                        eRubroMatricula = \
                                        Rubro.objects.filter(matricula=eMatricula, tipo_id=RUBRO_MATRICULA)[0]
                                        eRubroMatricula.relacionados = None
                                        eRubroMatricula.save()
                                        lista = []
                                        c = 0
                                        for r in ePeriodoMatricula.fecha_cuotas_rubro().values('fecha').distinct():
                                            c += 1
                                            lista.append([c, valor_cuota_mensual, r['fecha']])
                                        for item in lista:
                                            rubro = Rubro(tipo_id=RUBRO_ARANCEL,
                                                          persona=eMatricula.inscripcion.persona,
                                                          relacionados=eRubroMatricula,
                                                          matricula=eMatricula,
                                                          nombre=nombrearancel,
                                                          cuota=item[0],
                                                          fecha=datetime.now().date(),
                                                          fechavence=item[2],
                                                          valor=item[1],
                                                          iva_id=1,
                                                          valoriva=0,
                                                          valortotal=item[1],
                                                          saldo=item[1],
                                                          cancelado=False)
                                            rubro.save()
                                        arancel.delete()
                                        if ePeriodoMatricula.valida_rubro_acta_compromiso:
                                            isResult, message = eMatricula.generar_actacompromiso_matricula_pregrado()
                                            if not isResult:
                                                raise NameError(message)
                                            eMatricula.aranceldiferido = 1
                                            eMatricula.actacompromiso = message
                                            eMatricula.save()
                                eMatricula.calcula_nivel()
                                itinerario = 0
                                if not eMatricula.inscripcion.itinerario is None and eMatricula.inscripcion.itinerario > 0:
                                    itinerario = eMatricula.inscripcion.itinerario
                                cantidad_seleccionadas = 0
                                cursor = connections['default'].cursor()
                                sql = "select am.nivelmalla_id, count(am.nivelmalla_id) as cantidad_materias_seleccionadas from sga_materiaasignada ma, sga_materia m, sga_asignaturamalla am where ma.status=true and ma.matricula_id=" + str(
                                    eMatricula.id) + " and m.status=true and m.id=ma.materia_id and am.status=true and am.id=m.asignaturamalla_id GROUP by am.nivelmalla_id, am.malla_id order by count(am.nivelmalla_id) desc, am.nivelmalla_id desc limit 1;"
                                if itinerario > 0:
                                    sql = "select am.nivelmalla_id, count(am.nivelmalla_id) as cantidad_materias_seleccionadas from sga_materiaasignada ma, sga_materia m, sga_asignaturamalla am where ma.status=true and ma.matricula_id=" + str(
                                        eMatricula.id) + " and m.status=true and m.id=ma.materia_id and am.status=true and am.id=m.asignaturamalla_id and (am.itinerario=0 or am.itinerario=" + str(
                                        itinerario) + ") GROUP by am.nivelmalla_id, am.malla_id order by count(am.nivelmalla_id) desc, am.nivelmalla_id desc limit 1;"
                                cursor.execute(sql)
                                results = cursor.fetchall()
                                nivel = 0
                                for per in results:
                                    nivel = per[0]
                                    cantidad_seleccionadas = per[1]
                                cantidad_nivel = 0
                                eAsignaturaMallas = AsignaturaMalla.objects.filter(nivelmalla__id=nivel, status=True,
                                                                                   malla=eMatricula.inscripcion.mi_malla())
                                if itinerario > 0:
                                    eAsignaturaMallas = eAsignaturaMallas.filter(
                                        Q(itinerario=0) | Q(itinerario=itinerario))

                                for eAsignaturaMalla in eAsignaturaMallas:
                                    if Materia.objects.filter(nivel__periodo=eMatricula.nivel.periodo,
                                                              asignaturamalla=eAsignaturaMalla).exists():
                                        if eMatricula.inscripcion.estado_asignatura(eAsignaturaMalla.asignatura) != 1:
                                            cantidad_nivel += 1

                                porcentaje_seleccionadas = int(round(Decimal((float(cantidad_nivel) * float(
                                    PORCIENTO_PERDIDA_PARCIAL_GRATUIDAD)) / 100).quantize(Decimal('.00')), 0))
                                if (cantidad_seleccionadas < porcentaje_seleccionadas):
                                    tipo_matricula_ri = 2
                                else:
                                    tipo_matricula_ri = 1
                                eMatricula.grupo_socio_economico(tipo_matricula_ri)
                                eMatriculaGrupoSocioEconomicos = eMatricula.matriculagruposocioeconomico_set.all()
                                if eMatriculaGrupoSocioEconomicos.values('id').exists():
                                    eMatriculaGrupoSocioEconomico = eMatriculaGrupoSocioEconomicos[0]
                                    eMatriculaGrupoSocioEconomico.estado_gratuidad = eMatricula.inscripcion.estado_gratuidad
                                    eMatriculaGrupoSocioEconomico.save()
                        except Exception as ex:
                            print(f"Error al guardar en matricula: {eMatricula.__str__()} - Error: {ex.__str__()}")
                            transaction.set_rollback(True)
                            erorres.append(eInscripcion.pk)
                else:
                    break
    if len(erorres) > 0:
        print(f"Ocurrio errores de inscripciones: {erorres.__str__()}")
    if total_reajuste > 0:
        print(f"Se reajustaron {total_reajuste} inscripciones")


# eliminar_registro_record_materia(57387)


# ePeriodo = Periodo.objects.get(pk=153)
# ajuste_recalcular_rubros(ePeriodo)


# _eTitulos = Titulos('0927576504')
# eTitulos = _eTitulos.consultar()
# url = "http://181.224.196.52/api/dinardap/senescyt"
# auth = HTTPBasicAuth('senescyt', 'Robertito123**')
# headers = {"Content-Type": "application/json; charset=utf-8"}
#
# data = {
#     "identificacion": '0927576504',
# }
#
# # response = requests.post(url, headers=headers, json=data, auth=auth)
# response = requests.post(url, headers=headers, json=data, auth=auth)
# status = response.status_code
# if status == 200:
#     data = response.json()
# print(response)

def obtener_titulos_senescyt(ePeriodo):
    from soap.consumer.senescyt import Titulos
    from bd.models import HistorialPersonaConsultaTitulo
    eMatriculas = Matricula.objects.filter(status=True, nivel__periodo=ePeriodo)
    print(f"Total de matriculados {eMatriculas.count()}")
    eHistorialPersonaConsultaTitulos = HistorialPersonaConsultaTitulo.objects.filter(
        persona_id__in=eMatriculas.values_list('inscripcion__persona__id', flat=True)).distinct()
    print(f"Total de hisotiral {eHistorialPersonaConsultaTitulos.count()}")
    eMatriculas = eMatriculas.exclude(
        inscripcion__persona__id__in=eHistorialPersonaConsultaTitulos.values_list('persona__id', flat=True))
    total = eMatriculas.count()
    print(f"Total a procesar {total}")
    contador = 0
    for eMatricula in eMatriculas:
        try:
            contador += 1
            eInscripcion = eMatricula.inscripcion
            ePersona = eInscripcion.persona
            identificacion = ePersona.identificacion()
            eTitulos = Titulos(identificacion)
            mistitulos = eTitulos.consultar()
            print(f"{total}/{contador}")
            # eInscripcion.actualiza_estado_matricula()
            time.sleep(2)
        except Exception as ex:
            pass


# def obtener_titulos_senescyt(ePeriodo):
#     cursor = connections['admisionvirtual'].cursor()
#     sql = """
#
#     """
#     cursor.execute(sql)
#     results = cursor.fetchall()
#     nivel = 0
#     for per in results:
#         nivel = per[0]
#         cantidad_seleccionadas = per[1]
#     for eMatricula in eMatriculas:
#         try:
#             contador += 1
#             eInscripcion = eMatricula.inscripcion
#             ePersona = eInscripcion.persona
#             identificacion = ePersona.identificacion()
#             eTitulos = Titulos(identificacion)
#             mistitulos = eTitulos.consultar()
#             print(f"{total}/{contador}")
#             # eInscripcion.actualiza_estado_matricula()
#             time.sleep(2)
#         except Exception as ex:
#             pass


def activar_horario_examen_admision(ePeriodo):
    eMateriaAsignadaPlanificacionSedeVirtualExamenes = MateriaAsignadaPlanificacionSedeVirtualExamen.objects.filter(
        materiaasignada__matricula__nivel__periodo=ePeriodo, materiaasignada__matricula__inscripcion__coordinacion_id=9)
    total = len(eMateriaAsignadaPlanificacionSedeVirtualExamenes.values("id"))
    contador = 0
    for eMateriaAsignadaPlanificacionSedeVirtualExamen in eMateriaAsignadaPlanificacionSedeVirtualExamenes:
        contador += 1
        eMateriaAsignadaPlanificacionSedeVirtualExamen.utilizar_qr = True
        eMateriaAsignadaPlanificacionSedeVirtualExamen.save()
        eMateriaAsignada = eMateriaAsignadaPlanificacionSedeVirtualExamen.materiaasignada
        eMateriaAsignada.visiblehorarioexamen = True
        eMateriaAsignada.save()
        print(f"{total}/{contador} => {eMateriaAsignadaPlanificacionSedeVirtualExamen.__str__()}")


def partir_torta_admision(ePeriodo):
    aula_bloque_v_id = 284
    aula_crai_id = 47
    if not DEBUG:
        aula_bloque_v_id = 285
        aula_crai_id = 47
    eLaboratorioVirtual_1 = LaboratorioVirtual.objects.get(pk=aula_crai_id)
    eLaboratorioVirtual_2 = LaboratorioVirtual.objects.get(pk=aula_bloque_v_id)
    filtro_1 = Q(aulaplanificacion__turnoplanificacion__fechaplanificacion__sede_id=1) & Q(
        materiaasignada__matricula__inscripcion__coordinacion_id=9) & Q(
        materiaasignada__matricula__nivel__periodo=ePeriodo)
    eMateriaAsignadaPlanificacionSedeVirtualExamenes = MateriaAsignadaPlanificacionSedeVirtualExamen.objects.filter(
        filtro_1)
    eFechaPlanificacionSedeVirtualExamenes = FechaPlanificacionSedeVirtualExamen.objects.filter(sede_id=1,
                                                                                                periodo=ePeriodo,
                                                                                                pk__in=eMateriaAsignadaPlanificacionSedeVirtualExamenes.values_list(
                                                                                                    'aulaplanificacion__turnoplanificacion__fechaplanificacion__id',
                                                                                                    flat=True))
    for eFechaPlanificacionSedeVirtualExamen in eFechaPlanificacionSedeVirtualExamenes:
        print(f"Fecha: {eFechaPlanificacionSedeVirtualExamen}")
        eTurnoPlanificacionSedeVirtualExamenes = TurnoPlanificacionSedeVirtualExamen.objects.filter(
            fechaplanificacion=eFechaPlanificacionSedeVirtualExamen)
        for eTurnoPlanificacionSedeVirtualExamen in eTurnoPlanificacionSedeVirtualExamenes:
            filtro = filtro_1 & Q(aulaplanificacion__turnoplanificacion=eTurnoPlanificacionSedeVirtualExamen)
            eAulaPlanificacionSedeVirtualExamen = AulaPlanificacionSedeVirtualExamen.objects.get(
                turnoplanificacion=eTurnoPlanificacionSedeVirtualExamen, aula=eLaboratorioVirtual_1)
            MateriaAsignadaPlanificacionSedeVirtualExamen.objects.filter(
                aulaplanificacion__turnoplanificacion__fechaplanificacion__sede_id=1,
                materiaasignada__matricula__inscripcion__coordinacion_id=9,
                materiaasignada__matricula__nivel__periodo=ePeriodo,
                aulaplanificacion__turnoplanificacion=eTurnoPlanificacionSedeVirtualExamen).update(
                aulaplanificacion=eAulaPlanificacionSedeVirtualExamen)
            eMateriaAsignadaPlanificacionSedeVirtualExamenes = MateriaAsignadaPlanificacionSedeVirtualExamen.objects.filter(
                filtro)
            total = len(eMateriaAsignadaPlanificacionSedeVirtualExamenes.values("id"))
            print("Total: ", total)
            if total > 0:
                ids_1 = [x.id for x in
                         eMateriaAsignadaPlanificacionSedeVirtualExamenes[:eLaboratorioVirtual_2.capacidad]]
                eMateriaAsignadaPlanificacionSedeVirtualExamenes_1 = MateriaAsignadaPlanificacionSedeVirtualExamen.objects.filter(
                    pk__in=ids_1)
                print("Total BLOQUE V: ", len(eMateriaAsignadaPlanificacionSedeVirtualExamenes_1.values("id")))
                eMateriaAsignadaPlanificacionSedeVirtualExamenes_2 = eMateriaAsignadaPlanificacionSedeVirtualExamenes.exclude(
                    pk__in=ids_1)
                print("Total CRAI: ", len(eMateriaAsignadaPlanificacionSedeVirtualExamenes_2.values("id")))
                # CRAI
                eAulaPlanificacionSedeVirtualExamenes_1 = AulaPlanificacionSedeVirtualExamen.objects.filter(
                    turnoplanificacion=eTurnoPlanificacionSedeVirtualExamen, aula=eLaboratorioVirtual_1)
                if not eAulaPlanificacionSedeVirtualExamenes_1.values("id").exists():
                    eAulaPlanificacionSedeVirtualExamen_1 = AulaPlanificacionSedeVirtualExamen(
                        turnoplanificacion=eTurnoPlanificacionSedeVirtualExamen,
                        aula=eLaboratorioVirtual_1)
                    eAulaPlanificacionSedeVirtualExamen_1.save()
                else:
                    eAulaPlanificacionSedeVirtualExamen_1 = eAulaPlanificacionSedeVirtualExamenes_1.first()

                # BLOQUE_V
                eAulaPlanificacionSedeVirtualExamenes_2 = AulaPlanificacionSedeVirtualExamen.objects.filter(
                    turnoplanificacion=eTurnoPlanificacionSedeVirtualExamen, aula=eLaboratorioVirtual_2)
                if not eAulaPlanificacionSedeVirtualExamenes_2.values("id").exists():
                    eAulaPlanificacionSedeVirtualExamen_2 = AulaPlanificacionSedeVirtualExamen(
                        turnoplanificacion=eTurnoPlanificacionSedeVirtualExamen,
                        aula=eLaboratorioVirtual_2)
                    eAulaPlanificacionSedeVirtualExamen_2.save()
                else:
                    eAulaPlanificacionSedeVirtualExamen_2 = eAulaPlanificacionSedeVirtualExamenes_2.first()
                print(f"Se procesa BLOQUE V")
                eMateriaAsignadaPlanificacionSedeVirtualExamenes_1.update(
                    aulaplanificacion=eAulaPlanificacionSedeVirtualExamen_2, asistencia=False, fecha_asistencia=None)
                print(f"Se procesa CRAI")
                eMateriaAsignadaPlanificacionSedeVirtualExamenes_2.update(
                    aulaplanificacion=eAulaPlanificacionSedeVirtualExamen_1, asistencia=False, fecha_asistencia=None)


def generar_pdf_codigo_qr(ePeriodo):
    filtro_1 = Q(aulaplanificacion__turnoplanificacion__fechaplanificacion__sede_id=1) & Q(
        materiaasignada__matricula__inscripcion__coordinacion_id=9) & Q(
        materiaasignada__matricula__nivel__periodo=ePeriodo)
    eMateriaAsignadaPlanificacionSedeVirtualExamenes = MateriaAsignadaPlanificacionSedeVirtualExamen.objects.filter(
        filtro_1)
    eFechaPlanificacionSedeVirtualExamenes = FechaPlanificacionSedeVirtualExamen.objects.filter(sede_id=1,
                                                                                                periodo=ePeriodo,
                                                                                                pk__in=eMateriaAsignadaPlanificacionSedeVirtualExamenes.values_list(
                                                                                                    'aulaplanificacion__turnoplanificacion__fechaplanificacion__id',
                                                                                                    flat=True))
    for eFechaPlanificacionSedeVirtualExamen in eFechaPlanificacionSedeVirtualExamenes:
        print(f"Fecha: {eFechaPlanificacionSedeVirtualExamen}")
        eTurnoPlanificacionSedeVirtualExamenes = TurnoPlanificacionSedeVirtualExamen.objects.filter(
            fechaplanificacion=eFechaPlanificacionSedeVirtualExamen)
        for eTurnoPlanificacionSedeVirtualExamen in eTurnoPlanificacionSedeVirtualExamenes:
            eAulaPlanificacionSedeVirtualExamenes = AulaPlanificacionSedeVirtualExamen.objects.filter(
                turnoplanificacion=eTurnoPlanificacionSedeVirtualExamen)
            for eAulaPlanificacionSedeVirtualExamen in eAulaPlanificacionSedeVirtualExamenes:
                eMateriaAsignadaPlanificacionSedeVirtualExamenes = MateriaAsignadaPlanificacionSedeVirtualExamen.objects.filter(
                    aulaplanificacion=eAulaPlanificacionSedeVirtualExamen)
                for eMateriaAsignadaPlanificacionSedeVirtualExamen in eMateriaAsignadaPlanificacionSedeVirtualExamenes:
                    try:
                        eMatriculaSedeExamen = MatriculaSedeExamen.objects.get(
                            matricula=eMateriaAsignadaPlanificacionSedeVirtualExamen.materiaasignada.matricula)
                    except ObjectDoesNotExist:
                        eMatriculaSedeExamen = None
                    aceptotermino = False
                    genero_qr = False
                    url_pdf_examen = ''
                    if eMateriaAsignadaPlanificacionSedeVirtualExamen.utilizar_qr:
                        if eMatriculaSedeExamen:
                            if eMatriculaSedeExamen.aceptotermino:
                                aceptotermino = True
                                try:
                                    result = generate_qr_examen_final(eMateriaAsignadaPlanificacionSedeVirtualExamen,
                                                                      materiaasignada_id=eMateriaAsignadaPlanificacionSedeVirtualExamen.materiaasignada_id)
                                    isSuccess = result.get('isSuccess', False)
                                    if not isSuccess:
                                        message = result.get('message', 'Error al generar documento')
                                        raise NameError(message)
                                    aDataExamen = result.get('data', {})
                                    url_pdf_examen = aDataExamen.get('url_pdf', '')
                                    codigo_qr_examen = aDataExamen.get('codigo_qr', '')
                                    if url_pdf_examen == '' and codigo_qr_examen == '':
                                        raise NameError(u"No se encontro url del documento")
                                    eMateriaAsignadaPlanificacionSedeVirtualExamen.fecha_qr = datetime.now()
                                    eMateriaAsignadaPlanificacionSedeVirtualExamen.url_qr = f"/media/{url_pdf_examen}"
                                    eMateriaAsignadaPlanificacionSedeVirtualExamen.codigo_qr = codigo_qr_examen
                                    genero_qr = True
                                except Exception as ex:
                                    eMateriaAsignadaPlanificacionSedeVirtualExamen.fecha_qr = None
                                    eMateriaAsignadaPlanificacionSedeVirtualExamen.url_qr = None
                                    eMateriaAsignadaPlanificacionSedeVirtualExamen.codigo_qr = None
                                    genero_qr = False
                    if aceptotermino:
                        ePersona = eMateriaAsignadaPlanificacionSedeVirtualExamen.materiaasignada.matricula.inscripcion.persona
                        ePerfilUsuario = ePersona.perfilusuario_set.filter(status=True,
                                                                           inscripcion=eMateriaAsignadaPlanificacionSedeVirtualExamen.materiaasignada.matricula.inscripcion).first()
                        fecha = eFechaPlanificacionSedeVirtualExamen.fecha
                        horainicio = eTurnoPlanificacionSedeVirtualExamen.horainicio
                        fecha_visible = datetime.combine(fecha, horainicio)
                        titulo = 'Actualización de horario de examen'
                        cuerpo = f'La asignatura {eMateriaAsignadaPlanificacionSedeVirtualExamen.materiaasignada.materia.asignatura.nombre} se cambio de ubicación'
                        url = f"https://sgaestudiante.unemi.edu.ec/alu_documentos/examenes"
                        if genero_qr:
                            url = f"https://sga.unemi.edu.ec/media/{url_pdf_examen}"
                        eNotificacion = Notificacion(titulo=titulo,
                                                     cuerpo=cuerpo,
                                                     app_label='SIE',
                                                     destinatario=eMateriaAsignadaPlanificacionSedeVirtualExamen.materiaasignada.matricula.inscripcion.persona,
                                                     perfil=ePerfilUsuario,
                                                     leido=False,
                                                     visible=True,
                                                     url=url,
                                                     fecha_hora_visible=fecha_visible)
                        eNotificacion.save()
                    eMateriaAsignadaPlanificacionSedeVirtualExamen.save()
                    print(f"Se ajusta {eMateriaAsignadaPlanificacionSedeVirtualExamen.__str__()}")


def partir_torta_pregrado_unemi(ePeriodo):
    fecha = datetime(2023, 8, 7, 0, 0).date()
    aula_bloque_v_id = 285
    aula_crai_id = 47
    idh = 2764
    idf = 2764
    eLaboratorioVirtual_1 = LaboratorioVirtual.objects.get(pk=aula_crai_id)
    eLaboratorioVirtual_2 = LaboratorioVirtual.objects.get(pk=aula_bloque_v_id)
    filtro_1 = Q(aulaplanificacion__turnoplanificacion__fechaplanificacion__sede_id=1) & Q(
        materiaasignada__matricula__nivel__periodo=ePeriodo)
    filtro_1 = filtro_1 & Q(aulaplanificacion__turnoplanificacion__fechaplanificacion__fecha__gte=fecha)
    exclude = Q(materiaasignada__matricula__inscripcion__coordinacion_id=9)
    eMateriaAsignadaPlanificacionSedeVirtualExamenes = MateriaAsignadaPlanificacionSedeVirtualExamen.objects.filter(
        filtro_1).exclude(exclude)
    eFechaPlanificacionSedeVirtualExamenes = FechaPlanificacionSedeVirtualExamen.objects.filter(sede_id=1,
                                                                                                periodo=ePeriodo,
                                                                                                pk__in=eMateriaAsignadaPlanificacionSedeVirtualExamenes.values_list(
                                                                                                    'aulaplanificacion__turnoplanificacion__fechaplanificacion__id',
                                                                                                    flat=True))
    for eFechaPlanificacionSedeVirtualExamen in eFechaPlanificacionSedeVirtualExamenes:
        print(f"Fecha: {eFechaPlanificacionSedeVirtualExamen}")
        eTurnoPlanificacionSedeVirtualExamenes = TurnoPlanificacionSedeVirtualExamen.objects.filter(
            fechaplanificacion=eFechaPlanificacionSedeVirtualExamen).exclude(pk=idh)
        for eTurnoPlanificacionSedeVirtualExamen in eTurnoPlanificacionSedeVirtualExamenes:
            print(f"Turno: {eTurnoPlanificacionSedeVirtualExamen}")
            filtro = filtro_1 & Q(aulaplanificacion__turnoplanificacion=eTurnoPlanificacionSedeVirtualExamen)
            try:
                eAulaPlanificacionSedeVirtualExamen = AulaPlanificacionSedeVirtualExamen.objects.get(
                    turnoplanificacion=eTurnoPlanificacionSedeVirtualExamen, aula=eLaboratorioVirtual_1)
            except ObjectDoesNotExist:
                eAulaPlanificacionSedeVirtualExamen = None
            if eAulaPlanificacionSedeVirtualExamen:
                MateriaAsignadaPlanificacionSedeVirtualExamen.objects.filter(
                    aulaplanificacion__turnoplanificacion__fechaplanificacion__sede_id=1,
                    materiaasignada__matricula__nivel__periodo=ePeriodo,
                    aulaplanificacion__turnoplanificacion=eTurnoPlanificacionSedeVirtualExamen).exclude(
                    materiaasignada__matricula__inscripcion__coordinacion_id=9).update(
                    aulaplanificacion=eAulaPlanificacionSedeVirtualExamen)
                eMateriaAsignadaPlanificacionSedeVirtualExamenes = MateriaAsignadaPlanificacionSedeVirtualExamen.objects.filter(
                    filtro)
                total = len(eMateriaAsignadaPlanificacionSedeVirtualExamenes.values("id"))
                print("Total: ", total)
                if total > 0:
                    ids_1 = [x.id for x in
                             eMateriaAsignadaPlanificacionSedeVirtualExamenes[:eLaboratorioVirtual_2.capacidad]]
                    eMateriaAsignadaPlanificacionSedeVirtualExamenes_1 = MateriaAsignadaPlanificacionSedeVirtualExamen.objects.filter(
                        pk__in=ids_1)
                    print("Total BLOQUE V: ", len(eMateriaAsignadaPlanificacionSedeVirtualExamenes_1.values("id")))
                    eMateriaAsignadaPlanificacionSedeVirtualExamenes_2 = eMateriaAsignadaPlanificacionSedeVirtualExamenes.exclude(
                        pk__in=ids_1)
                    print("Total CRAI: ", len(eMateriaAsignadaPlanificacionSedeVirtualExamenes_2.values("id")))
                    # CRAI
                    eAulaPlanificacionSedeVirtualExamenes_1 = AulaPlanificacionSedeVirtualExamen.objects.filter(
                        turnoplanificacion=eTurnoPlanificacionSedeVirtualExamen, aula=eLaboratorioVirtual_1)
                    if not eAulaPlanificacionSedeVirtualExamenes_1.values("id").exists():
                        eAulaPlanificacionSedeVirtualExamen_1 = AulaPlanificacionSedeVirtualExamen(
                            turnoplanificacion=eTurnoPlanificacionSedeVirtualExamen,
                            aula=eLaboratorioVirtual_1)
                        eAulaPlanificacionSedeVirtualExamen_1.save()
                    else:
                        eAulaPlanificacionSedeVirtualExamen_1 = eAulaPlanificacionSedeVirtualExamenes_1.first()

                    # BLOQUE_V
                    eAulaPlanificacionSedeVirtualExamenes_2 = AulaPlanificacionSedeVirtualExamen.objects.filter(
                        turnoplanificacion=eTurnoPlanificacionSedeVirtualExamen, aula=eLaboratorioVirtual_2)
                    if not eAulaPlanificacionSedeVirtualExamenes_2.values("id").exists():
                        eAulaPlanificacionSedeVirtualExamen_2 = AulaPlanificacionSedeVirtualExamen(
                            turnoplanificacion=eTurnoPlanificacionSedeVirtualExamen,
                            aula=eLaboratorioVirtual_2)
                        eAulaPlanificacionSedeVirtualExamen_2.save()
                    else:
                        eAulaPlanificacionSedeVirtualExamen_2 = eAulaPlanificacionSedeVirtualExamenes_2.first()
                    print(f"Se procesa BLOQUE V")
                    eMateriaAsignadaPlanificacionSedeVirtualExamenes_1.update(
                        aulaplanificacion=eAulaPlanificacionSedeVirtualExamen_2, asistencia=False,
                        fecha_asistencia=None, utilizar_qr=True, fecha_qr=None, url_qr=None, codigo_qr=None)
                    print(f"Se procesa CRAI")
                    eMateriaAsignadaPlanificacionSedeVirtualExamenes_2.update(
                        aulaplanificacion=eAulaPlanificacionSedeVirtualExamen_1, asistencia=False,
                        fecha_asistencia=None, utilizar_qr=True)


def notificar_cambio_ubicacion_pregrado(ePeriodo):
    fecha = datetime(2023, 8, 7, 0, 0).date()
    aula_bloque_v_id = 285
    aula_crai_id = 47
    idh = 2764
    idf = 2764
    filtro_1 = Q(aulaplanificacion__turnoplanificacion__fechaplanificacion__sede_id=1) & Q(
        materiaasignada__matricula__nivel__periodo=ePeriodo)
    filtro_1 = filtro_1 & Q(aulaplanificacion__turnoplanificacion__fechaplanificacion__fecha__gte=fecha)
    exclude = Q(materiaasignada__matricula__inscripcion__coordinacion_id=9)
    eMateriaAsignadaPlanificacionSedeVirtualExamenes = MateriaAsignadaPlanificacionSedeVirtualExamen.objects.filter(
        filtro_1).exclude(exclude)
    eFechaPlanificacionSedeVirtualExamenes = FechaPlanificacionSedeVirtualExamen.objects.filter(sede_id=1,
                                                                                                periodo=ePeriodo,
                                                                                                pk__in=eMateriaAsignadaPlanificacionSedeVirtualExamenes.values_list(
                                                                                                    'aulaplanificacion__turnoplanificacion__fechaplanificacion__id',
                                                                                                    flat=True))
    for eFechaPlanificacionSedeVirtualExamen in eFechaPlanificacionSedeVirtualExamenes:
        print(f"Fecha: {eFechaPlanificacionSedeVirtualExamen}")
        eTurnoPlanificacionSedeVirtualExamenes = TurnoPlanificacionSedeVirtualExamen.objects.filter(
            fechaplanificacion=eFechaPlanificacionSedeVirtualExamen).exclude(pk=idh)
        for eTurnoPlanificacionSedeVirtualExamen in eTurnoPlanificacionSedeVirtualExamenes:
            print(f"Turno: {eTurnoPlanificacionSedeVirtualExamen}")
            eAulaPlanificacionSedeVirtualExamenes = AulaPlanificacionSedeVirtualExamen.objects.filter(
                turnoplanificacion=eTurnoPlanificacionSedeVirtualExamen)
            for eAulaPlanificacionSedeVirtualExamen in eAulaPlanificacionSedeVirtualExamenes:
                eMateriaAsignadaPlanificacionSedeVirtualExamenes = MateriaAsignadaPlanificacionSedeVirtualExamen.objects.filter(
                    aulaplanificacion=eAulaPlanificacionSedeVirtualExamen)
                for eMateriaAsignadaPlanificacionSedeVirtualExamen in eMateriaAsignadaPlanificacionSedeVirtualExamenes:
                    if eMateriaAsignadaPlanificacionSedeVirtualExamen.aulaplanificacion.aula_id == aula_bloque_v_id:
                        try:
                            eMatriculaSedeExamen = MatriculaSedeExamen.objects.get(
                                matricula=eMateriaAsignadaPlanificacionSedeVirtualExamen.materiaasignada.matricula)
                        except ObjectDoesNotExist:
                            eMatriculaSedeExamen = None
                        aceptotermino = False
                        genero_qr = False
                        url_pdf_examen = ''
                        if eMateriaAsignadaPlanificacionSedeVirtualExamen.utilizar_qr:
                            if eMatriculaSedeExamen:
                                if eMatriculaSedeExamen.aceptotermino:
                                    aceptotermino = True
                                    # try:
                                    #     result = generate_qr_examen_final(eMateriaAsignadaPlanificacionSedeVirtualExamen, materiaasignada_id=eMateriaAsignadaPlanificacionSedeVirtualExamen.materiaasignada_id)
                                    #     isSuccess = result.get('isSuccess', False)
                                    #     if not isSuccess:
                                    #         message = result.get('message', 'Error al generar documento')
                                    #         raise NameError(message)
                                    #     aDataExamen = result.get('data', {})
                                    #     url_pdf_examen = aDataExamen.get('url_pdf', '')
                                    #     codigo_qr_examen = aDataExamen.get('codigo_qr', '')
                                    #     if url_pdf_examen == '' and codigo_qr_examen == '':
                                    #         raise NameError(u"No se encontro url del documento")
                                    #     eMateriaAsignadaPlanificacionSedeVirtualExamen.fecha_qr = datetime.now()
                                    #     eMateriaAsignadaPlanificacionSedeVirtualExamen.url_qr = f"/media/{url_pdf_examen}"
                                    #     eMateriaAsignadaPlanificacionSedeVirtualExamen.codigo_qr = codigo_qr_examen
                                    #     genero_qr = True
                                    # except Exception as ex:
                                    #     eMateriaAsignadaPlanificacionSedeVirtualExamen.fecha_qr = None
                                    #     eMateriaAsignadaPlanificacionSedeVirtualExamen.url_qr = None
                                    #     eMateriaAsignadaPlanificacionSedeVirtualExamen.codigo_qr = None
                                    #     genero_qr = False
                                    # eMateriaAsignadaPlanificacionSedeVirtualExamen.fecha_qr = None
                                    # eMateriaAsignadaPlanificacionSedeVirtualExamen.url_qr = None
                                    # eMateriaAsignadaPlanificacionSedeVirtualExamen.codigo_qr = None
                        if aceptotermino:
                            ePersona = eMateriaAsignadaPlanificacionSedeVirtualExamen.materiaasignada.matricula.inscripcion.persona
                            ePerfilUsuario = ePersona.perfilusuario_set.filter(status=True,
                                                                               inscripcion=eMateriaAsignadaPlanificacionSedeVirtualExamen.materiaasignada.matricula.inscripcion).first()
                            fecha = eFechaPlanificacionSedeVirtualExamen.fecha
                            horainicio = eTurnoPlanificacionSedeVirtualExamen.horainicio
                            fecha_visible = datetime.combine(fecha, horainicio)
                            titulo = 'Actualización de horario de examen'
                            cuerpo = f'La asignatura {eMateriaAsignadaPlanificacionSedeVirtualExamen.materiaasignada.materia.asignatura.nombre} se cambio de ubicación, vuelva a generar el código QR'
                            url = f"https://sgaestudiante.unemi.edu.ec/alu_documentos/examenes"
                            # if genero_qr:
                            #     url = f"https://sga.unemi.edu.ec/media/{url_pdf_examen}"
                            eNotificacion = Notificacion(titulo=titulo,
                                                         cuerpo=cuerpo,
                                                         app_label='SIE',
                                                         destinatario=eMateriaAsignadaPlanificacionSedeVirtualExamen.materiaasignada.matricula.inscripcion.persona,
                                                         perfil=ePerfilUsuario,
                                                         leido=False,
                                                         visible=True,
                                                         url=url,
                                                         fecha_hora_visible=fecha_visible)
                            eNotificacion.save()
                        # eMateriaAsignadaPlanificacionSedeVirtualExamen.save()
                        print(f"Se ajusta {eMateriaAsignadaPlanificacionSedeVirtualExamen.__str__()}")


def acivar_examenes_moodle(ePeriodo):
    fecha = datetime(2023, 8, 7, 0, 0).date()
    filtro_1 = Q(materiaasignada__matricula__nivel__periodo=ePeriodo) & Q(
        aulaplanificacion__turnoplanificacion__fechaplanificacion__fecha__gte=fecha)
    eMateriaAsignadaPlanificacionSedeVirtualExamenes = MateriaAsignadaPlanificacionSedeVirtualExamen.objects.filter(
        filtro_1)
    eFechaPlanificacionSedeVirtualExamenes = FechaPlanificacionSedeVirtualExamen.objects.filter(periodo=ePeriodo,
                                                                                                pk__in=eMateriaAsignadaPlanificacionSedeVirtualExamenes.values_list(
                                                                                                    'aulaplanificacion__turnoplanificacion__fechaplanificacion__id',
                                                                                                    flat=True))
    total_fechas = eFechaPlanificacionSedeVirtualExamenes.values("id").count()
    contador_fecha = 0
    for eFechaPlanificacionSedeVirtualExamen in eFechaPlanificacionSedeVirtualExamenes:
        contador_fecha += 1
        print(f"({total_fechas}/{contador_fecha}) Fecha: {eFechaPlanificacionSedeVirtualExamen.fecha}")
        eTurnoPlanificacionSedeVirtualExamenes = TurnoPlanificacionSedeVirtualExamen.objects.filter(
            fechaplanificacion=eFechaPlanificacionSedeVirtualExamen)
        total_turnos = eTurnoPlanificacionSedeVirtualExamenes.values("id").count()
        contador_turno = 0
        for eTurnoPlanificacionSedeVirtualExamen in eTurnoPlanificacionSedeVirtualExamenes:
            contador_turno += 1
            print(
                f"({total_turnos}/{contador_turno}) Turno: {eTurnoPlanificacionSedeVirtualExamen.horainicio} - {eTurnoPlanificacionSedeVirtualExamen.horafin}")
            eAulaPlanificacionSedeVirtualExamenes = AulaPlanificacionSedeVirtualExamen.objects.filter(
                turnoplanificacion=eTurnoPlanificacionSedeVirtualExamen)
            total_aulas = eAulaPlanificacionSedeVirtualExamenes.values("id").count()
            contador_aula = 0
            for eAulaPlanificacionSedeVirtualExamen in eAulaPlanificacionSedeVirtualExamenes:
                contador_aula += 1
                print(f"({total_aulas}/{contador_aula}) Aula: {eAulaPlanificacionSedeVirtualExamen.aula.nombre}")
                with transaction.atomic():
                    try:
                        # password = eAulaPlanificacionSedeVirtualExamen.password
                        # if not eAulaPlanificacionSedeVirtualExamen.password:
                        eAulaPlanificacionSedeVirtualExamen.password = None
                        eAulaPlanificacionSedeVirtualExamen.create_update_password()
                        eAulaPlanificacionSedeVirtualExamen.save()
                        password = eAulaPlanificacionSedeVirtualExamen.password
                    except Exception as ex:
                        transaction.set_rollback(True)
                        print(f"Error al generar password")
                fecha = eFechaPlanificacionSedeVirtualExamen.fecha
                horainicio = eTurnoPlanificacionSedeVirtualExamen.horainicio
                horafin = eTurnoPlanificacionSedeVirtualExamen.horafin
                fechadesde = datetime(fecha.year, fecha.month, fecha.day, horainicio.hour, horainicio.minute,
                                      horainicio.second)
                fechadesde = int(time.mktime(fechadesde.timetuple()))
                fechahasta = datetime(fecha.year, fecha.month, fecha.day, horafin.hour, horafin.minute, horafin.second)
                fechahasta = int(time.mktime(fechahasta.timetuple()))
                eMateriaAsignadaPlanificacionSedeVirtualExamenes = MateriaAsignadaPlanificacionSedeVirtualExamen.objects.filter(
                    aulaplanificacion=eAulaPlanificacionSedeVirtualExamen)
                total_alumnos = eMateriaAsignadaPlanificacionSedeVirtualExamenes.values("id").count()
                contador_alumno = 0
                habilitado = False
                for eMateriaAsignadaPlanificacionSedeVirtualExamen in eMateriaAsignadaPlanificacionSedeVirtualExamenes:
                    contador_alumno += 1
                    with transaction.atomic():
                        try:
                            eMateriaAsignada = eMateriaAsignadaPlanificacionSedeVirtualExamen.materiaasignada
                            username = eMateriaAsignada.matricula.inscripcion.persona.usuario.username
                            examenplanificado = eMateriaAsignada.materia.examenplanificadosilabo(
                                eMateriaAsignadaPlanificacionSedeVirtualExamen.detallemodeloevaluativo)
                            enviado = False
                            if examenplanificado:
                                quiz = buscarQuiz(examenplanificado.get("idtestmoodle"),
                                                  eMateriaAsignada.materia.coordinacion().id)
                                if quiz:
                                    limite = int(quiz[3])
                                    intentos = 1
                                    estado_examen = estadoQuizIndividual(username, eMateriaAsignada.materia,
                                                                         examenplanificado.get("idtestmoodle"))
                                    if estado_examen != 'inprogress':
                                        eMateriaAsignadaPlanificacionSedeVirtualExamen.idtestmoodle = int(
                                            examenplanificado.get("idtestmoodle"))
                                        eMateriaAsignadaPlanificacionSedeVirtualExamen.password = password
                                        eMateriaAsignada = eMateriaAsignadaPlanificacionSedeVirtualExamen.materiaasignada
                                        isResult, msgError = accesoQuizIndividual(
                                            eMateriaAsignada.matricula.inscripcion.persona.usuario.username,
                                            eMateriaAsignada.materia,
                                            eMateriaAsignadaPlanificacionSedeVirtualExamen.idtestmoodle,
                                            {'timeopen': fechadesde,
                                             'timeclose': fechahasta,
                                             'timelimit': limite,
                                             'password': eMateriaAsignadaPlanificacionSedeVirtualExamen.password,
                                             'attempts': intentos})
                                        if isResult:
                                            eMateriaAsignadaPlanificacionSedeVirtualExamen.habilitadoexamen = True
                                            enviado = True
                                            habilitado = True
                                        eMateriaAsignadaPlanificacionSedeVirtualExamen.save()

                            if enviado:
                                print(
                                    f"({total_alumnos}/{contador_alumno}) Habilitado alumno: {eMateriaAsignadaPlanificacionSedeVirtualExamen}")
                            else:
                                print(
                                    f"({total_alumnos}/{contador_alumno}) No habilitado alumno: {eMateriaAsignadaPlanificacionSedeVirtualExamen}")
                        except Exception as ex:
                            transaction.set_rollback(True)
                            print(f"Error {ex.__str__()}")
                if habilitado:
                    with transaction.atomic():
                        try:
                            eAulaPlanificacionSedeVirtualExamen.registrohabilitacion = datetime.now()
                            eAulaPlanificacionSedeVirtualExamen.save()
                        except Exception as ex:
                            transaction.set_rollback(True)
                            print(f"Error al habilitar aula")


def reporte_examen_pregrado_en_linea(ePeriodo):
    try:
        libre_origen = '/reporte_examen_sede_moodle_pregrado.xls'
        fuentecabecera = easyxf('font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
        fuentenormal = easyxf('font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
        # output_folder = MEDIA_ROOT
        output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
        # output_folder = os.path.join(os.path.join(BASE_DIR))
        # liborigen = xlrd.open_workbook(output_folder + libre_origen)
        libdestino = xlwt.Workbook()
        hojadestino = libdestino.add_sheet('HOJA1')
        fil = 0
        columnas = [
            (u"facultad", 7000, 1),
            (u"carrera", 7000, 1),
            (u"nivel", 7000, 0),
            (u"paralelo", 7000, 0),
            (u"asignatura", 7000, 0),
            (u"docente", 7000, 0),
            (u"idmateria", 7000, 0),
            (u"idcursomoodle", 7000, 0),
            (u"examen_creado_sga", 7000, 0),
            (u"modelo_evaluativo_sga", 7000, 0),
            (u"migrado_moodle", 7000, 0),
            (u"nombre_examen_moodle", 7000, 0),
            (u"desde_examen_moodle", 7000, 0),
            (u"hasta_examen_moodle", 7000, 0),
            (u"tiempo_examen_moodle", 7000, 0),
            (u"metodo_navegacion_examen_moodle", 7000, 0),
            (u"total_examen_moodle", 7000, 0),
            (u"valor_examen_moodle", 7000, 0),
            (u"tiene_clave_examen_moodle", 7000, 0),
            (u"categoria_calificacion_examen_moodle", 7000, 0),
            (u"item_calificacion_examen_moodle", 7000, 0),
            (u"seccion_examen_moodle", 7000, 0),
            (u"tiene_preguntas", 7000, 0)
        ]
        for col_num in range(len(columnas)):
            hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
            hojadestino.col(col_num).width = columnas[col_num][1]
        asignaturas_id = DetalleGrupoAsignatura.objects.filter(status=True, grupo_id=2).values('asignatura_id')
        eMaterias = Materia.objects.filter(Q(asignaturamalla__malla__carrera__modalidad=3) | Q(asignaturamalla__asignatura__id__in=asignaturas_id), status=True, nivel__periodo=ePeriodo).exclude(nivel__id__in=[1516, 1517, 1565, 1566, 1550, 1556])
        totalmaterias = eMaterias.count()
        cont = 1
        fila = 1
        for eMateria in eMaterias:
            facultad = eMateria.coordinacion()
            carrera = eMateria.asignaturamalla.malla.carrera
            nivel = eMateria.asignaturamalla.nivelmalla
            paralelo = eMateria.paralelo
            asignatura = eMateria.asignaturamalla.asignatura.nombre
            profesor = 'sin profesor'
            # pm = eMateria.profesormateria_set.filter(status=True, tipoprofesor_id=14).last()
            # if pm:
            #     profesor = pm.profesor.persona.nombre_completo_inverso()
            if (eProfesor := eMateria.profesor_principal()) is not None:
                profesor = eProfesor.persona.nombre_completo_inverso()
            idmateria = eMateria.id
            idcurso = eMateria.idcursomoodle
            tiene_examen_sga = 'NO'
            migrado_moodle = 'NO'
            num_preguntas = 0
            modeloevaluativo = 'SIN MODELO'
            if eMateria.modeloevaluativo:
                modeloevaluativo = eMateria.modeloevaluativo.nombre
            eTestSilaboSemanal = TestSilaboSemanal.objects.filter(status=True,
                                                                  silabosemanal__silabo__materia_id=idmateria,
                                                                  silabosemanal__examen=True,
                                                                  detallemodelo__alternativa_id=20).last()
            if facultad.id == 9:
                cursor_verbose = 'db_moodle_virtual'
            else:
                cursor_verbose = 'moodle_db'
            conexion = connections[cursor_verbose]
            cursor = conexion.cursor()
            name = ''
            timeopen = ''
            timeclose = ''
            timelimit = ''
            navmethod = ''
            sumgrades = ''
            grade = ''
            password = None
            categoria = eTestSilaboSemanal.detallemodelo.nombre if eTestSilaboSemanal else ''
            categoryid = 0
            categoria_moodle = ''
            itemid = 0
            item_moodle = ''
            seccion = ''
            instance = None
            if eTestSilaboSemanal:
                tiene_examen_sga = 'SI'
                if eTestSilaboSemanal.estado_id == 4:
                    migrado_moodle = 'SI'
                    instance = eTestSilaboSemanal.idtestmoodle
                if instance:
                    sql = f"""SELECT name, timeopen, timeclose, timelimit, navmethod, sumgrades, grade, password  FROM mooc_quiz WHERE id={instance} AND course={eMateria.idcursomoodle}"""
                    cursor.execute(sql)
                    quiz = cursor.fetchone()
                    if quiz:
                        name = quiz[0]
                        timeopen = quiz[1]
                        timeopen = str(datetime.fromtimestamp(timeopen))
                        timeclose = quiz[2]
                        timeclose = str(datetime.fromtimestamp(timeclose))
                        timelimit = quiz[3]
                        timelimit = str((timelimit / 60) if timelimit else 0)
                        navmethod = quiz[4]
                        sumgrades = str(quiz[5])
                        grade = str(quiz[6])
                        password = str(quiz[7])
                        sql = """SELECT id, fullname FROM mooc_grade_categories WHERE courseid=%s AND fullname='%s' and depth='2' """ % (
                        eMateria.idcursomoodle, categoria)
                        cursor.execute(sql)
                        category = cursor.fetchone()
                        if category:
                            categoryid = category[0]
                            categoria_moodle = category[1]
                            sql = """select id, itemname from mooc_grade_items WHERE courseid=%s AND categoryid=%s and itemname='%s' and iteminstance=%s """ % (
                            eMateria.idcursomoodle, categoryid, name, instance)
                            cursor.execute(sql)
                            item = cursor.fetchone()
                            if item:
                                itemid = item[0]
                                item_moodle = item[1]
                        sql = """SELECT section FROM mooc_course_modules WHERE course=%s AND instance=%s """ % (
                        eMateria.idcursomoodle, instance)
                        cursor.execute(sql)
                        course_module = cursor.fetchone()
                        if course_module:
                            sectionid = course_module[0]
                            sql = """SELECT name FROM mooc_course_sections WHERE course=%s AND id=%s """ % (
                            eMateria.idcursomoodle, sectionid)
                            cursor.execute(sql)
                            section = cursor.fetchone()
                            if section:
                                seccion = section[0]
                        sql = """SELECT DISTINCT  qet.name, qet.questiontext, re.answer, re.answerformat FROM 
                        mooc_quiz q INNER JOIN mooc_quiz_slots qe ON q.id=qe.quizid INNER JOIN mooc_question qet ON 
                        qet.category=qe.questioncategoryid INNER JOIN mooc_question_answers re ON re.question=qet.id 
                        WHERE re.fraction>0 AND q.id=%s  """ % (instance)
                        cursor.execute(sql)
                        preguntas = cursor.fetchall()
                        num_preguntas = len(preguntas)

            print('curso actualizado', cont, 'de', totalmaterias)
            cont += 1

            hojadestino.write(fila, 0, "%s" % facultad, fuentenormal)
            hojadestino.write(fila, 1, "%s" % carrera, fuentenormal)
            hojadestino.write(fila, 2, "%s" % nivel, fuentenormal)
            hojadestino.write(fila, 3, "%s" % paralelo, fuentenormal)
            hojadestino.write(fila, 4, "%s" % asignatura, fuentenormal)
            hojadestino.write(fila, 5, "%s" % profesor, fuentenormal)
            hojadestino.write(fila, 6, idmateria, fuentenormal)
            hojadestino.write(fila, 7, idcurso, fuentenormal)
            hojadestino.write(fila, 8, tiene_examen_sga, fuentenormal)
            hojadestino.write(fila, 9, modeloevaluativo, fuentenormal)
            hojadestino.write(fila, 10, migrado_moodle, fuentenormal)
            hojadestino.write(fila, 11, name, fuentenormal)
            hojadestino.write(fila, 12, timeopen, fuentenormal)
            hojadestino.write(fila, 13, timeclose, fuentenormal)
            hojadestino.write(fila, 14, timelimit, fuentenormal)
            hojadestino.write(fila, 15, navmethod, fuentenormal)
            hojadestino.write(fila, 16, sumgrades, fuentenormal)
            hojadestino.write(fila, 17, grade, fuentenormal)
            hojadestino.write(fila, 18, 'SI' if password else 'NO', fuentenormal)
            hojadestino.write(fila, 19, categoria_moodle, fuentenormal)
            hojadestino.write(fila, 20, item_moodle, fuentenormal)
            hojadestino.write(fila, 21, seccion, fuentenormal)
            hojadestino.write(fila, 22, 'SI' if num_preguntas > 0 else 'NO', fuentenormal)

            fila += 1

        libdestino.save(output_folder + libre_origen)
        print(output_folder + libre_origen)
        print("Proceso finalizado. . .")

    except Exception as ex:
        msg = ex.__str__()
        textoerror = '{} Linea:{}'.format(str(ex), sys.exc_info()[-1].tb_lineno)
        print(textoerror)
        print(msg)


def obtener_examen_transversales_pregrado(ePeriodo):
    print(f"INICIA")
    try:
        no_silabos = []
        asignaturas_id = DetalleGrupoAsignatura.objects.filter(status=True)
        filtro = Q(status=True, nivel__periodo=ePeriodo)
        filtro_1 = Q(Q(asignaturamalla__malla__carrera__modalidad__in=[1, 2]) & Q(
            asignaturamalla__asignatura__id__in=asignaturas_id.values('asignatura_id').filter(grupo_id=2)))
        filtro_2 = Q(Q(asignaturamalla__malla__carrera__modalidad__in=[3]) & Q(
            asignaturamalla__asignatura__id__in=asignaturas_id.values('asignatura_id')))
        eMaterias = Materia.objects.filter(Q(Q(filtro_1) | Q(filtro_2)) & Q(filtro)).exclude(
            nivel__id__in=[1481, 1482, 1501, 1508])
        total_materias = eMaterias.count()
        contador_materia = 0
        detallemodelo_id = 123
        semana = 32
        numsemana = 16
        for eMateria in eMaterias:
            contador_materia += 1
            eTestSilaboSemanal = None
            with transaction.atomic():
                try:
                    eSilabo = Silabo.objects.filter(materia=eMateria, status=True, codigoqr=True).last()
                    if eSilabo is None:
                        print(f"Materia no tiene silabo {eMateria.__str__()}")
                        no_silabos.append(eMateria.pk)
                    else:
                        try:
                            eSilaboSemanal = SilaboSemanal.objects.get(silabo=eSilabo, numsemana=numsemana,
                                                                       semana=semana, examen=True)
                        except ObjectDoesNotExist:
                            eSilaboSemanal = SilaboSemanal(silabo=eSilabo,
                                                           numsemana=numsemana,
                                                           semana=semana,
                                                           fechainiciosemana=datetime(2023, 8, 7, 0, 0).date(),
                                                           fechafinciosemana=datetime(2023, 8, 20, 0, 0).date(),
                                                           objetivoaprendizaje='',
                                                           enfoque='',
                                                           recursos='',
                                                           evaluacion='',
                                                           estado=3,
                                                           estadocumplimiento=2,
                                                           examen=True
                                                           )
                            eSilaboSemanal.save()
                        # instruccion = """La honestidad es el pilar fundamental de toda sociedad, por ello la UNEMI motiva el desarrollo de este valor en cada uno de sus estudiantes.&nbsp;<br><br>Lea cada una de las preguntas y responda.&nbsp;<br><br>Esta actividad representa el 60% de la nota final.&nbsp;<br><br><br>"""
                        instruccion = ''
                        recomendacion = ''
                        nombretest = 'EXAMEN FINAL'
                        fechadesde = datetime(2023, 8, 7, 0, 0, 1)
                        horadesde = datetime(2023, 8, 7, 0, 0, 1).time()
                        fechahasta = datetime(2023, 8, 31, 0, 0, 1)
                        horahasta = datetime(2023, 8, 31, 23, 59, 59).time()
                        vecesintento = 1
                        tiempoduracion = 60
                        calificar = True
                        navegacion = 1
                        migrado = True
                        password = '123JKDS@4pl'
                        try:
                            eTestSilaboSemanal = TestSilaboSemanal.objects.get(status=True,
                                                                               silabosemanal=eSilaboSemanal,
                                                                               detallemodelo_id=detallemodelo_id)
                        except ObjectDoesNotExist:
                            eTestSilaboSemanal = TestSilaboSemanal(detallemodelo_id=detallemodelo_id,
                                                                   silabosemanal=eSilaboSemanal,
                                                                   tiporecurso_id=11,
                                                                   estado_id=4,
                                                                   )
                        eTestSilaboSemanal.detallemodelo_id = detallemodelo_id
                        eTestSilaboSemanal.tiporecurso_id = 11
                        eTestSilaboSemanal.estado_id = 4
                        eTestSilaboSemanal.instruccion = instruccion
                        eTestSilaboSemanal.recomendacion = recomendacion
                        eTestSilaboSemanal.fechadesde = fechadesde
                        eTestSilaboSemanal.horadesde = horadesde
                        eTestSilaboSemanal.fechahasta = fechahasta
                        eTestSilaboSemanal.horahasta = horahasta
                        eTestSilaboSemanal.vecesintento = vecesintento
                        eTestSilaboSemanal.tiempoduracion = tiempoduracion
                        eTestSilaboSemanal.calificar = calificar
                        eTestSilaboSemanal.navegacion = navegacion
                        eTestSilaboSemanal.migrado = migrado
                        eTestSilaboSemanal.password = password
                        eTestSilaboSemanal.nombretest = nombretest
                        eTestSilaboSemanal.save()

                except Exception as ex:
                    transaction.set_rollback(True)
                    print(f"Error {ex.__str__()}")

            eCoordinacion = eMateria.coordinacion()
            if eCoordinacion.id == 9:
                cursor_verbose = 'db_moodle_virtual'
            else:
                cursor_verbose = 'moodle_db'

            conexion = connections[cursor_verbose]
            cursor = conexion.cursor()
            idcursomoodle = eMateria.idcursomoodle
            iteminstance = 0
            url = None
            if eTestSilaboSemanal:
                categoria = eTestSilaboSemanal.detallemodelo.nombre if eTestSilaboSemanal else ''
                sql = """SELECT id, fullname FROM mooc_grade_categories WHERE courseid=%s AND fullname='%s' and depth='2' """ % (
                idcursomoodle, categoria)
                cursor.execute(sql)
                category = cursor.fetchone()
                if category:
                    categoryid = category[0]
                    categoria_moodle = category[1]
                    sql = """SELECT id, iteminstance FROM mooc_grade_items WHERE courseid=%s AND categoryid=%s AND itemtype='mod' AND itemmodule='quiz' """ % (
                    idcursomoodle, categoryid)
                    cursor.execute(sql)
                    item = cursor.fetchone()
                    if item:
                        itemid = item[0]
                        iteminstance = item[1]
                        sql = """SELECT id FROM mooc_course_modules WHERE course=%s AND instance=%s """ % (
                        idcursomoodle, iteminstance)
                        cursor.execute(sql)
                        modules = cursor.fetchone()
                        if modules:
                            url = f'https://aulagrado.unemi.edu.ec/mod/quiz/view.php?id={modules[0]}'
            with transaction.atomic():
                try:
                    eTestSilaboSemanal.idtestmoodle = iteminstance
                    eTestSilaboSemanal.url = url
                    eTestSilaboSemanal.save()
                    print(f"Se actualizo test de moodle en Materia: {eMateria.__str__()}")
                except Exception as ex:
                    transaction.set_rollback(True)
                    print(f"Error {ex.__str__()}")
            print(f"{total_materias}/{contador_materia} Materia: {eMateria.__str__()}")
        if len(no_silabos):
            print(f"Materias sin silabos: {no_silabos.__str__()}")

    except Exception as ex:
        msg = ex.__str__()
        textoerror = '{} Linea:{}'.format(str(ex), sys.exc_info()[-1].tb_lineno)
        print(textoerror)
        print(msg)
    print(f"FIN")


# reporte_examen_pregrado_en_linea(Periodo.objects.get(pk=177))
# obtener_examen_transversales_pregrado(Periodo.objects.get(pk=177))
# acivar_examenes_moodle(Periodo.objects.get(pk=177))


def crear_lecciones(profesor, fecha, clases):
    puede_crear_clases = True
    for cl in clases:
        if cl.tipoprofesor.id in [2, 13] and cl.tipohorario in [1, 2, 8]:
            puede_crear_clases = False
            break
    if not puede_crear_clases:
        return False, "No existe clases a crear"
    with transaction.atomic():
        try:
            if LeccionGrupo.objects.values("id").filter(profesor=profesor, turno=clases[0].turno, fecha=fecha).exists():
                lecciongrupo = LeccionGrupo.objects.get(profesor=profesor, turno=clases[0].turno, fecha=fecha)
                lecciongrupo.status = True
                lecciongrupo.abierta = False
                lecciongrupo.turno = clases[0].turno
                lecciongrupo.aula = clases[0].aula
                lecciongrupo.dia = clases[0].dia
                lecciongrupo.fecha = fecha
                lecciongrupo.horaentrada = clases[0].turno.comienza
                lecciongrupo.horasalida = clases[0].turno.termina
                lecciongrupo.save()
            else:
                lecciongrupo = LeccionGrupo(profesor=profesor,
                                            turno=clases[0].turno,
                                            aula=clases[0].aula,
                                            dia=clases[0].dia,
                                            fecha=fecha,
                                            horaentrada=clases[0].turno.comienza,
                                            horasalida=clases[0].turno.termina,
                                            abierta=False,
                                            contenido='SIN CONTENIDO',
                                            estrategiasmetodologicas='SIN CONTENIDO',
                                            observaciones='SIN OBSERVACIONES',
                                            status=True
                                            )
                lecciongrupo.save()
            for cl in clases:
                if Leccion.objects.values("id").filter(clase=cl, fecha=fecha).exists():
                    leccion = Leccion.objects.get(clase=cl, fecha=fecha)
                    leccion.abierta = False
                    leccion.status = True
                    leccion.horaentrada = cl.turno.comienza
                    leccion.horasalida = cl.turno.termina
                    leccion.contenido = lecciongrupo.contenido
                    leccion.observaciones = lecciongrupo.observaciones
                    leccion.save()
                else:
                    leccion = Leccion(clase=cl,
                                      fecha=fecha,
                                      horaentrada=cl.turno.comienza,
                                      horasalida=cl.turno.termina,
                                      abierta=False,
                                      status=True,
                                      contenido=lecciongrupo.contenido,
                                      observaciones=lecciongrupo.observaciones
                                      )
                    leccion.save()
                if not lecciongrupo.lecciones.values("id").filter(pk=leccion.id).exists():
                    lecciongrupo.lecciones.add(leccion)
                materia = cl.materia
                asignados = materia.asignados_a_esta_materia()
                for materiaasignada in asignados:
                    if AsistenciaLeccion.objects.values("id").filter(leccion=leccion,
                                                                     materiaasignada=materiaasignada).exists():
                        asistencialeccion = \
                        AsistenciaLeccion.objects.filter(leccion=leccion, materiaasignada=materiaasignada)[0]
                        asistencialeccion.leccion = leccion
                        asistencialeccion.materiaasignada = materiaasignada
                        asistencialeccion.asistio = True
                        asistencialeccion.virtual = False
                        asistencialeccion.virtual_fecha = None
                        asistencialeccion.virtual_hora = None
                        asistencialeccion.ip_private = None
                        asistencialeccion.ip_public = None
                        asistencialeccion.browser = None
                        asistencialeccion.ops = None
                        asistencialeccion.screen_size = None
                        asistencialeccion.status = True
                    else:
                        asistencialeccion = AsistenciaLeccion(leccion=leccion,
                                                              materiaasignada=materiaasignada,
                                                              asistio=True,
                                                              virtual=False,
                                                              virtual_fecha=None,
                                                              virtual_hora=None,
                                                              ip_private=None,
                                                              ip_public=None,
                                                              browser=None,
                                                              ops=None,
                                                              screen_size=None,
                                                              status=True,
                                                              )
                    asistencialeccion.save()
                    asistencialeccion
            lecciongrupo.save()
            return True, "Lección creada"
        except Exception as ex:
            transaction.set_rollback(True)
            return False, f"*** Error {ex.__str__()}"


def crear_lecciones_dia_especifico(periodo, fecha):
    # fecha = date(2024, 1, 2, 0, 0)
    numerosemana = fecha.isocalendar()[1]
    dia_semana = fecha.isoweekday()
    print(u"****************************************************************************************************")
    print(
        f"Inicia proceso a las {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} de crear lecciones del día {DIAS_CHOICES[dia_semana - 1][1]} del periodo {periodo}")
    fechas = [fecha]
    ePeriodoAcademia = periodo.get_periodoacademia()
    for fecha in fechas:
        dia = (fecha.weekday() + 1)
        profesores = Profesor.objects.filter(profesormateria__materia__nivel__periodo=periodo,
                                             profesormateria__activo=True).distinct()
        # if DEBUG:
        #     profesores = profesores.filter(profesormateria__profesor_id__in=[3170])
        total_profesores = len(profesores)
        print("*** FECHA A PROCESAR: " + fecha.__str__() + "\r")
        contP = 1
        for profesor in profesores:
            numClass = 0
            print(f"***** ({contP}/{total_profesores}) -> Profesor: {profesor.__str__()}")
            contP += 1
            clases = Clase.objects.filter(activo=True, inicio__lte=fecha, fin__gte=fecha, dia=dia, status=True,
                                          materia__nivel__periodo=periodo,
                                          materia__nivel__periodo__visible=True,
                                          materia__nivel__periodo__visiblehorario=True,
                                          materia__fechafinasistencias__gte=fecha,
                                          materia__profesormateria__profesor_id=profesor.id,
                                          profesor_id=profesor.id,
                                          materia__profesormateria__tipoprofesor_id__in=[1, 2, 5, 7, 8, 10, 11, 12, 13,
                                                                                         14, 15],
                                          tipoprofesor_id__in=[1, 2, 5, 7, 8, 10, 11, 12, 13, 14, 15])
            # clases = clases.filter(turno__comienza__gte=time(6, 0, 0), turno__termina__lte=time(23, 59, 00))
            clases = clases.distinct()
            clases_turnos = profesor.extraer_clases_y_turnos_practica(fecha, periodo)
            sesiones = Sesion.objects.filter(Q(turno__id__in=clases.values_list('turno__id').distinct())).distinct()
            clases = clases.filter(
                Q(pk__in=clases.values_list('id')) | Q(pk__in=clases_turnos[0].values_list('id'))).distinct()
            for sesion in sesiones:
                for turno in sesion.turnos_clasehorario(clases):
                    aux_clasesactuales = turno.horario_profesor_actual_horario(dia, profesor, periodo)
                    total_clases = len(aux_clasesactuales)
                    for cl in aux_clasesactuales:
                        contC = 1
                        print(f"******** ({contC}/{total_clases}) -> Clase: {cl.__str__()}")
                        fechacompara = cl.compararfecha(numerosemana)
                        # tieneferiado = periodo.es_feriado(fechacompara, cl.materia)
                        coordinacion = cl.materia.coordinacion()
                        modalidad = cl.materia.asignaturamalla.malla.modalidad
                        # if not tieneferiado:
                        if coordinacion.id in [1, 2, 3, 4, 5, 12, 9]:
                            if cl.tipohorario == 1:
                                if aux_clasesactuales[0].tipoprofesor.id != 8:
                                    isSuccess, msg = crear_lecciones(profesor, fecha,
                                                                     turno.horario_profesor_actual_horario(dia,
                                                                                                           profesor,
                                                                                                           periodo,
                                                                                                           False,
                                                                                                           False))
                                    if isSuccess:
                                        contC += 1
                                        numClass += 1
                                    else:
                                        print(msg)
                                # 2 => CLASE VIRTUAL SINCRÓNICA
                                # 8 => CLASE REFUERZO SINCRÓNICA
                                # 7 => CLASE VIRTUAL ASINCRÓNICA
                                # 9 => CLASE REFUERZO ASINCRÓNICA
                            elif cl.tipohorario in [2, 7, 8, 9]:
                                if modalidad:
                                    if modalidad.id in [1, 2]:
                                        if cl.tipohorario in [2, 8]:
                                            isSuccess, msg = crear_lecciones(profesor, fecha,
                                                                             turno.horario_profesor_actual_horario(dia,
                                                                                                                   profesor,
                                                                                                                   periodo,
                                                                                                                   False,
                                                                                                                   False))
                                            if isSuccess:
                                                contC += 1
                                                numClass += 1
                                            else:
                                                print(msg)
                                    elif modalidad.id in [3]:
                                        if cl.tipohorario in [2, 8]:
                                            isSuccess, msg = crear_lecciones(profesor, fecha,
                                                                             turno.horario_profesor_actual_horario(dia,
                                                                                                                   profesor,
                                                                                                                   periodo,
                                                                                                                   False,
                                                                                                                   False))
                                            if isSuccess:
                                                contC += 1
                                                numClass += 1
                                            else:
                                                print(msg)
                                        elif cl.tipohorario in [7, 9]:
                                            if cl.subirenlace:
                                                clasesactualesasincronica = cl.horario_profesor_actualasincronica(
                                                    numerosemana)
                                                if not clasesactualesasincronica.values("id").exists():
                                                    isSuccess, msg = crear_lecciones(profesor, fecha,
                                                                                     turno.horario_profesor_actual_horario(
                                                                                         dia, profesor, periodo, False,
                                                                                         False))
                                                    if isSuccess:
                                                        contC += 1
                                                        numClass += 1
                                                    else:
                                                        print(msg)

            print(f"***** (Se crearon {numClass} lecciones) -> Profesor: {profesor.__str__()}")


# fecha = datetime(2024, 1, 2, 0, 0)
# crear_lecciones_dia_especifico(Periodo.objects.get(pk=224), fecha)


def reporte_derecho_en_linea(ePeriodo):
    import os
    import random
    import xlwt
    from settings import MEDIA_ROOT, MEDIA_URL, DEBUG
    from sga.models import Matricula, Malla, RequisitoTitulacionMalla, Persona, Notificacion, Periodo
    from webpush import send_user_notification
    _ePersona = Persona.objects.get(pk=10730)
    ePeriodo = Periodo.objects.get(pk=317)
    try:
        directory = os.path.join(MEDIA_ROOT, 'reportes')
        try:
            os.stat(directory)
        except:
            os.mkdir(directory)
        nombre_archivo = "reporte_derecho_en_linea_{}.xls".format(random.randint(1, 10000).__str__())
        directory = os.path.join(MEDIA_ROOT, 'reportes', nombre_archivo)
        fuentecabecera = xlwt.easyxf('font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
        fuentenormal = xlwt.easyxf('font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
        wb = xlwt.Workbook()
        ws = wb.add_sheet("Hoja1")
        fil = 0
        columnas = [
            (u"idMatricula", 7000, 0),
            (u"Documento", 7000, 0),
            (u"NombreCompleto", 7000, 0),
            (u"NivelMatricula", 7000, 0),
            (u"idMateria", 7000, 0),
            (u"Asignatura", 7000, 0),
            (u"NivelAsignatura", 7000, 0),
        ]
        aRequisitos = []
        eMalla = Malla.objects.get(pk=480)
        eRequisitoTitulacionMallas = RequisitoTitulacionMalla.objects.filter(malla=eMalla, status=True).order_by('requisito__nombre')
        num = 1
        for eRequisitoTitulacionMalla in eRequisitoTitulacionMallas:
            columnas.append((f"R{num}", 7000, 0))
            aRequisitos.append({'id': eRequisitoTitulacionMalla.id, 'name': eRequisitoTitulacionMalla.requisito.nombre, 'num': num})
        for col_num in range(len(columnas)):
            ws.write(fil, col_num, columnas[col_num][0], fuentecabecera)
            ws.col(col_num).width = columnas[col_num][1]
        filtro = Q(nivel__periodo=ePeriodo) & Q(status=True) & Q(inscripcion__inscripcionmalla__malla=eMalla) & \
                 Q(inscripcion__inscripcionmalla__status=True)
        eMatriculas = Matricula.objects.filter(filtro)
        total = eMatriculas.values("id").count()
        cont = 1
        fila = 1
        for eMatricula in eMatriculas[:10]:
            eInscripcion = eMatricula.inscripcion
            ePersona = eInscripcion.persona
            _eMaterias = eMatricula.materias()
            for _eMateria in _eMaterias:
                eMateria = _eMateria.materia
                print(f"({total}/{cont}) -> {ePersona.nombre_completo_inverso()} >> {eMateria.asignaturamalla.asignatura.nombre}")
                ws.write(fila, 0, "%s" % eMatricula.id, fuentenormal)  # idmatricula
                ws.write(fila, 1, "%s" % ePersona.documento(), fuentenormal)  # documento
                ws.write(fila, 2, "%s" % ePersona.nombre_completo_inverso(), fuentenormal)  # nombre_completo
                ws.write(fila, 3, "%s" % eMatricula.nivelmalla.nombre if eMatricula.nivelmalla else "", fuentenormal)  # NivelMatricula
                ws.write(fila, 4, "%s" % eMateria.id, fuentenormal) #idMateria
                ws.write(fila, 5, "%s" % eMateria.asignaturamalla.asignatura.nombre, fuentenormal) #Asignatura
                ws.write(fila, 6, "%s" % eMateria.asignaturamalla.nivelmalla.nombre, fuentenormal) #NivelAsignatura
                col = 6
                if len(eRequisitos := eMateria.requisitomateriaunidadintegracioncurricular_set.filter(status=True).order_by('requisito__nombre')) > 0:
                    for aRequisito in aRequisitos:
                        col += 1
                        idRequisito = aRequisito.get('id', '0')
                        eRequisito = eRequisitos.filter(requisito_id=idRequisito).first()
                        if eRequisito:
                            cumple = eRequisito.run(eInscripcion.pk)
                            ws.write(fila, col, "%s" % 'SI' if cumple else 'NO', fuentenormal)  #REQUISITO
                        else:
                            ws.write(fila, col, "-", fuentenormal)  #REQUISITO
                fila += 1
                cont += 1

        ws.save(directory)

        eNotificacion = Notificacion(cuerpo=f'Reporte de estudiantes de la carrera de derecho en línea listo',
                                     titulo=f'Reporte de estudiantes de la carrera de derecho en línea',
                                     destinatario=_ePersona,
                                     url="{}reportes/{}".format(MEDIA_URL, nombre_archivo),
                                     prioridad=1,
                                     app_label='SGA',
                                     fecha_hora_visible=datetime.now() + timedelta(days=1),
                                     tipo=2,
                                     en_proceso=False)
        eNotificacion.save()
        send_user_notification(user=_ePersona.usuario,
                               payload={"head": "Reporte de estudiantes de la carrera de derecho en línea",
                                        "body": 'Reporte de estudiantes de la carrera de derecho en línea listo',
                                        "action": "notificacion",
                                        "timestamp": time.mktime(datetime.now().timetuple()),
                                        "url": "{}reportes/{}".format(MEDIA_URL, nombre_archivo),
                                        "btn_notificaciones": traerNotificaciones(None, None, _ePersona),
                                        "mensaje": 'Su reporte ha sido generado con exito'
                                        },
                               ttl=500)
        print("Proceso finalizado. . .")

    except Exception as ex:
        msg = ex.__str__()
        textoerror = '{} Linea:{}'.format(str(ex), sys.exc_info()[-1].tb_lineno)
        print(textoerror)
        print(msg)
        eNotificacion = Notificacion(cuerpo=textoerror,
                                     titulo=f'Reporte de estudiantes de la carrera de derecho en línea ha fallado',
                                     destinatario=_ePersona,
                                     prioridad=1,
                                     app_label='SGA',
                                     fecha_hora_visible=datetime.now() + timedelta(days=1),
                                     tipo=2,
                                     en_proceso=False,
                                     error=True)
        eNotificacion.save()
        send_user_notification(user=_ePersona.usuario,
                               payload={"head": "Reporte de estudiantes de la carrera de derecho en línea ha fallado en la ejecución",
                                        "body": 'Reporte de estudiantes de la carrera de derecho en línea ha fallado',
                                        "action": "notificacion",
                                        "timestamp": time.mktime(datetime.now().timetuple()),
                                        "btn_notificaciones": traerNotificaciones(None, None, _ePersona),
                                        "mensaje": textoerror,
                                        "error": True
                                        },
                               ttl=500)



reporte_derecho_en_linea(Periodo.objects.get(pk=317))
