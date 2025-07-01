import os
import sys
import io
import xlsxwriter
import xlwt
import openpyxl
import xlwt
from xlwt import *
from django.http import HttpResponse
from xlwt import *

# SITE_ROOT = os.path.dirname(os.path.realpath(__file__))
YOUR_PATH = os.path.dirname(os.path.realpath(__file__))
# print(f"YOUR_PATH: {YOUR_PATH}")
SITE_ROOT = os.path.dirname(os.path.dirname(YOUR_PATH))
SITE_ROOT = os.path.join(SITE_ROOT, '')
# print(f"SITE_ROOT: {SITE_ROOT}")
your_djangoproject_home = os.path.split(SITE_ROOT)[0]
# print(f"your_djangoproject_home: {your_djangoproject_home}")
sys.path.append(your_djangoproject_home)

from django.core.wsgi import get_wsgi_application

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "settings")
application = get_wsgi_application()

from django.http import HttpResponse
from settings import MEDIA_ROOT, BASE_DIR
from xlwt import easyxf, XFStyle
from sga.adm_criteriosactividadesdocente import asistencia_tutoria
from inno.models import *
from sga.models import *
from sagest.models import *
from balcon.models import *
from inno.funciones import *
from Moodle_Funciones import crearhtmlphpmoodle
from sga.funciones import log, convertir_fecha, puede_realizar_accion, puede_realizar_accion_afirmativo, \
    null_to_decimal, generar_nombre, fechatope, convertir_fecha_invertida, variable_valor, MiPaginador, \
    dia_semana_ennumero_fecha, null_to_numeric, calculate_username, generar_usuario,generar_usuario_admision
from settings import MATRICULACION_LIBRE, UTILIZA_GRUPOS_ALUMNOS, NOMBRE_NIVEL_AUTOMATICO, MATRICULACION_POR_NIVEL, \
    CAPACIDAD_MATERIA_INICIAL, CUPO_POR_MATERIA, APROBACION_DISTRIBUTIVO, USA_EVALUACION_INTEGRAL, \
    TIPO_DOCENTE_TEORIA, TIPO_DOCENTE_PRACTICA, VERIFICAR_CONFLICTO_DOCENTE, TIPO_CUOTA_RUBRO, SITE_STORAGE, \
    HORAS_VIGENCIA, ADMISION_ID, USA_TIPOS_INSCRIPCIONES, NOTA_ESTADO_EN_CURSO, TIPO_INSCRIPCION_INICIAL, \
    TIPO_DOCENTE_FIRMA, TIPO_DOCENTE_AYUDANTIA, EMAIL_INSTITUCIONAL_AUTOMATICO, EMAIL_DOMAIN, ALUMNOS_GROUP_ID
import concurrent.futures


def cerrar_materias_transversales_5():
    materias = Materia.objects.filter(status=True, nivel__periodo_id=177, cerrado=False,
                                      asignaturamalla__malla__carrera__coordinacion__in=[5], modeloevaluativo_id=27)
    for materia in materias:
        print(materia)
        for asig in materia.asignados_a_esta_materia():
            asig.cerrado = True
            asig.save(actualiza=False)
            asig.actualiza_estado()
        for asig in materia.asignados_a_esta_materia():
            asig.cierre_materia_asignada()

        materia.cerrado = True
        materia.fechacierre = datetime.now().date()
        materia.save()

#cerrar_materias_transversales_5()

def actualizar_nivel_inscripcion_malla8():
    matriculas = Matricula.objects.filter(status=True, nivel__periodo_id=177, inscripcion__carrera_id__in=[127,141,175] )
    for matricula in matriculas:
        inscripcion = matricula.inscripcion
        print('ACTUALIZANDO- ', inscripcion.persona.cedula)
        inscripcion.actualizar_nivel()
        print('ACTUALIZADO')
    print('FIN')

#actualizar_nivel_inscripcion_malla8()

def calificacion_transversales_en_linea():
    try:
        periodo = Periodo.objects.get(id=177)
        asignaturas = DetalleGrupoAsignatura.objects.values_list('asignatura_id', flat=True).filter(status=True,
                                                                                                    grupo_id__in=[1, 2,
                                                                                                                  3])

        materias = Materia.objects.filter(status=True, nivel__periodo_id=177,
                                          asignaturamalla__asignatura_id__in=asignaturas,
                                          modeloevaluativo_id=27, asignaturamalla__malla__carrera__coordinacion__in=[5])

        for materia in materias:
            idcursomoodle = materia.idcursomoodle
            materiasasignadas = MateriaAsignada.objects.filter(status=True,
                                                               matricula__nivel__periodo=periodo,
                                                               materia=materia,
                                                               matricula__bloqueomatricula=False,
                                                               matricula__retiradomatricula=False, materia__status=True,
                                                               matricula__status=True,
                                                               matricula__inscripcion__carrera__modalidad__in=[1, 2, 3])

            cursor = connections['moodle_db'].cursor()
            for materiaasignada in materiasasignadas:
                # guardo_nota = False
                usuario = materiaasignada.matricula.inscripcion.persona.usuario.username
                # SE NECESITA EL ID DE CURSO MOODLE
                sql = """
                                                    SELECT ROUND(nota.finalgrade,2), UPPER(gc.fullname)
                                                            FROM mooc_grade_grades nota
                                                    INNER JOIN mooc_grade_items it ON nota.itemid=it.id AND courseid=%s AND itemtype='category'
                                                    INNER JOIN mooc_grade_categories gc ON gc.courseid=it.courseid AND gc.id=it.iteminstance AND gc.depth=2
                                                    INNER JOIN mooc_user us ON nota.userid=us.id
                                                    WHERE us.username = '%s' and UPPER(gc.fullname)='RE'
                                                    ORDER BY it.sortorder
                                                """ % (str(idcursomoodle), usuario)

                cursor.execute(sql)
                results = cursor.fetchall()
                if results:
                    for notasmooc in results:
                        campo = materiaasignada.campo(notasmooc[1].upper())
                        if not campo:
                            print('revisar curso moodle - ', materiaasignada.materia.id, 'idcursomoodle -',
                                  materiaasignada.materia.idcursomoodle)
                            continue
                        if type(notasmooc[0]) is Decimal:
                            if null_to_decimal(campo.valor) != float(notasmooc[0]):
                                actualizar_nota_planificacion(materiaasignada.id, notasmooc[1].upper(), notasmooc[0])
                                auditorianotas = AuditoriaNotas(evaluaciongenerica=campo, manual=False,
                                                                calificacion=notasmooc[0])
                                auditorianotas.save()
                                print('importacion exitosa - ', materiaasignada)

                        else:
                            if null_to_decimal(campo.valor) != float(0):
                                actualizar_nota_planificacion(materiaasignada.id, notasmooc[1].upper(), notasmooc[0])
                                auditorianotas = AuditoriaNotas(evaluaciongenerica=campo, manual=False, calificacion=0)
                                auditorianotas.save()
                                print('importacion exitosa - ', materiaasignada)

                else:
                    detallemodelo = DetalleModeloEvaluativo.objects.get(pk=125)
                    campo = materiaasignada.campo(detallemodelo.nombre)
                    actualizar_nota_planificacion(materiaasignada.id, detallemodelo.nombre, 0)
                    auditorianotas = AuditoriaNotas(evaluaciongenerica=campo, manual=False, calificacion=0)
                    auditorianotas.save()

            for asig in materia.asignados_a_esta_materia():
                asig.cerrado = True
                asig.save(actualiza=False)
                asig.actualiza_estado()
                asig.cierre_materia_asignada()

            print('PROCESO FINALIZADO')






    except Exception as ex:
        msg = ex.__str__()

        textoerror = '{} Linea:{}'.format(str(ex), sys.exc_info()[-1].tb_lineno)
        print(textoerror)
        print(msg)


#calificacion_transversales_en_linea()

def actualizar_correos():
    with transaction.atomic():
        try:
            print("Inicio Actualizar correos")
            matriculados= Matricula.objects.filter(status=True, nivel__periodo_id=224, nivel__id__in=[1516,1517], termino=True, inscripcion__persona__emailinst='')
            c = 0
            cc= matriculados.count()
            for matricula in matriculados:
                persona= matricula.inscripcion.persona
                if persona.usuario is None:
                    username = calculate_username(persona)
                    usuario = generar_usuario_admision(persona, username, ALUMNOS_GROUP_ID)
                    persona.emailinst = username + '@' + EMAIL_DOMAIN
                    persona.save(usuario_id=persona.usuario.id)
                else:
                    persona.emailinst = persona.usuario.username + '@' + EMAIL_DOMAIN
                    persona.save(usuario_id=persona.usuario.id)
                c+=1
                print(f'{c} correos actualizados de {cc}')

            print(f'FIN! -  {c} correos actualizados')
        except Exception as ex:
            transaction.set_rollback(True)
            textoerror = '{} Linea:{} - Info: {} '.format(str(ex), sys.exc_info()[-1].tb_lineno)
            print(textoerror)


#actualizar_correos()

def activar_deudores_eliminados():
    cedulas = ['0302375126',
'0503775157',
'0503775157',
'0650342702',
'0650342702',
'0705550887',
'0705550887',
'0928055201',
'0928055201',
'0928278431',
'0928278431',
'0928322783',
'0928322783',
'0929216299',
'0931702989',
'0931702989',
'0941661530',
'0941661530',
'0942098351',
'0942098351',
'0950223909',
'0950223909',
'0953254471',
'1105035636',
'1105035636',
'1718254574',
'1718254574',
'1719613885',
'1719613885',
'1723915953',
'1723915953',
'1752899938',
'1752899938',
'1105428625',
'0706479680',
'1105428625',
'0925815375',
'0926927328',
'0926927328',
'0929824803',
'0929824803',
'0930397716',
'0930397716',
'0940238926',
'0940238926',
'0956245732',
'1309522744',
'1311576399',
'1311576399',
'1400512750',
'1400512750',
'1500747751',
'1500747751',
'1722563663',
'1722563663',
'1105314304',
'0959268533',
'1311903379',
'1317962072',
'0926048166',
'0925953366',
'0942554772',
'1720586435',
'0802830257',
'0202012746',
'0705948933',
'2450211988',
'0942098351',
'0950223909',
'1205428186',
'1105030934',
'0944405026'

               ]
    try:
        cnmoodle = connections['moodle_db'].cursor()
        matriculas= Matricula.objects.filter(nivel__periodo__id=224, inscripcion__persona__cedula__in=cedulas)
        for matricula in matriculas:
            matricula.status=True
            matricula.save()
            print('matricula activa')
            materiasasignadas = MateriaAsignada.objects.filter(status=False, matricula=matricula)
            for materia in materiasasignadas:
                materia.status=True
                materia.save()
                print('materia activa')
            rubros=Rubro.objects.filter(status=False, matricula=matricula)
            for rubro in rubros:
                rubro.status=True
                rubro.save()
                print('rubro activo')
            usermoodle = matricula.inscripcion.persona.usuario.username
            if usermoodle:
                sql = f"Select id, username From mooc_user Where username='{usermoodle}'"
                cnmoodle.execute(sql)
                registro = cnmoodle.fetchall()
                # idusuario = registro[0][0]
                try:
                    usermoodle = registro[0][1]
                    # Asignar estado suspended = 1 para que no pueda acceder al aula virtual
                    sql = f"Update mooc_user Set suspended=0 Where username='{usermoodle}'"
                    cnmoodle.execute(sql)
                except Exception as ex:
                    print(f'********--------{usermoodle}--------********')

    except Exception as ex:
        transaction.set_rollback(True)
        print('error: %s' % ex)

@transaction.atomic()
def enlaces_meet():
    try:
        miarchivo = openpyxl.load_workbook("docentes_enlaces.xlsx")
        lista = miarchivo.get_sheet_by_name('ENLACE MEET')
        totallista = lista.rows
        a = 0
        for filas in totallista:
            a += 1
            if a > 1:
                ceduladocente = str(filas[1].value)
                if Profesor.objects.filter(status=True, persona__cedula=ceduladocente).exists():
                    docente = Profesor.objects.get(status=True, persona__cedula=ceduladocente)
                    docente.urlzoom = str(filas[6].value)
                    docente.save()
                    print('enlace actualizado', ceduladocente)
                else:
                    print('no se encontró docente', ceduladocente)
        print('FIN')

    except Exception as ex:
        transaction.set_rollback(True)
        print('error: %s' % ex)

def homologacion_comunicacion_rezagados():

    #verificar que todos los estudiantes tengan la misma malla
    cadena = ''
    linea, excluidos, conexito = 0, 0, 0
    try:
        libre_origen = '/homologacion_comunicacion_rezagados.xls'
        fuentecabecera = easyxf(
            'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
        fuentenormal = easyxf(
            'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')

        output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
        # liborigen = xlrd.open_workbook(output_folder + libre_origen)
        libdestino = xlwt.Workbook()
        hojadestino = libdestino.add_sheet('Sheet1')
        fil = 0
        columnas = [(u"CEDULA", 6000),
                    (u"APELLIDOS Y NOMBRES", 6000),
                    (u"OBSERVACIÓN", 6000),
                    (u"HORAS PRACTICAS", 6000),
                    (u"HORAS VINCULACION", 6000),
                    (u"OBSERVACIÓN", 6000)
                    ]

        for col_num in range(len(columnas)):
            hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
            hojadestino.col(col_num).width = columnas[col_num][1]
        fila = 1
        titulacion = 0

        lin = 0
        #miarchivo = openpyxl.load_workbook("CEDULA_COMUNICACION_final.xlsx")
        miarchivo = openpyxl.load_workbook("inscripcion_comunicacion.xlsx")
        #miarchivo = openpyxl.load_workbook("prueba_2.xlsx")

        ws = miarchivo.get_sheet_by_name("prueba")
        worksheet = ws
        c = 0
        cont = 0
        periodo_id=224
        carrera_id=131
        mallaantigua_id=205
        mallanueva_id=488
        sin_matricula = []

        for row in worksheet.iter_rows(min_row=0):
            if lin >= 0:
                currentValues, cadena = [], ''
                for cell in row:
                    cadena += str(cell.value) + ' '
                    currentValues.append(str(cell.value))
                inscripcion = int(currentValues[0])

                if not inscripcion:
                    break

                with transaction.atomic():
                    matricula = Matricula.objects.filter(status=True, inscripcion__carrera__id=carrera_id,
                                                         inscripcion_id=inscripcion).first()
                    if matricula:
                        cont += 1
                        matricula.pasoayuda = True
                        matricula.save()
                        print(u"%s - %s" % (matricula, cont))
                        inscripcion = matricula.inscripcion
                        hojadestino.write(fila, 0, matricula.inscripcion.persona.identificacion(), fuentenormal)
                        hojadestino.write(fila, 1, matricula.inscripcion.persona.nombre_completo(), fuentenormal)
                        hojadestino.write(fila, 2, matricula.nivel.periodo.nombre, fuentenormal)

                        practicaspp= haber_cumplido_horas_creditos_practicas_preprofesionales(inscripcion.id)
                        horasvinculacion = haber_cumplido_horas_creditos_vinculacion(inscripcion.id)

                        if InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallaantigua_id).exists():
                            imantigua = InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallaantigua_id)[0]
                            imantigua.status = False
                            imantigua.save()
                            print(u"Desactiva antigua inscripcion -----------------------------")

                        if not InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallanueva_id).exists():
                            imnueva = InscripcionMalla(inscripcion=inscripcion, malla_id=mallanueva_id)
                            imnueva.save()
                            print(u"Crea nueva inscripcion -----------------------------")

                        equivalencias = TablaEquivalenciaAsignaturas.objects.filter(status=True, asignaturamalla__malla_id=mallaantigua_id).order_by('asignaturamallasalto__nivelmalla__orden')
                        cont_asig_vinculacion_aprobadas = 0
                        for equivalencia in equivalencias:
                            print(u"nueva - %s" % equivalencia.asignaturamallasalto)
                            recordantiguo = inscripcion.recordacademico_set.filter(status=True,asignaturamalla_id=equivalencia.asignaturamalla_id).first()

                            if recordantiguo:
                                print(u"anterior - %s" % equivalencia.asignaturamalla)
                                print(u"Record antiguo: %s" % recordantiguo)
                                recordnuevo = None
                                recordantiguo.status = False
                                recordantiguo.save(update_asignaturamalla=False)

                                if equivalencia.asignaturamallasalto_id in [10850,10853,10854,10859,10865]:
                                    observaciones = recordantiguo.observaciones + " Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15"
                                    homologada = True



                                else:
                                    observaciones = recordantiguo.observaciones + " Migración con base a RESOLUCIÓN CGA-SO-9-2023-NO15 Y RESOLUCIÓN OCS-SO-18-2023-NO14"
                                    homologada = False
                                if not RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                      asignaturamalla=equivalencia.asignaturamallasalto).exists():



                                    recordnuevo = RecordAcademico(inscripcion=inscripcion,
                                                                  matriculas=recordantiguo.matriculas,
                                                                  asignaturamalla=equivalencia.asignaturamallasalto,
                                                                  asignatura=equivalencia.asignaturamallasalto.asignatura,
                                                                  asignaturaold_id=recordantiguo.asignatura.id,
                                                                  nota=recordantiguo.nota,
                                                                  asistencia=recordantiguo.asistencia,
                                                                  sinasistencia=recordantiguo.sinasistencia,
                                                                  fecha=recordantiguo.fecha,
                                                                  noaplica=recordantiguo.noaplica,
                                                                  aprobada=recordantiguo.aprobada,
                                                                  convalidacion=recordantiguo.convalidacion,
                                                                  pendiente=recordantiguo.pendiente,
                                                                  creditos=equivalencia.asignaturamallasalto.creditos,
                                                                  horas=equivalencia.asignaturamallasalto.horas,
                                                                  valida=recordantiguo.valida,
                                                                  validapromedio=recordantiguo.validapromedio,
                                                                  observaciones=observaciones,
                                                                  homologada=homologada,
                                                                  materiaregular=recordantiguo.materiaregular,
                                                                  materiacurso=None,
                                                                  completonota=recordantiguo.completonota,
                                                                  completoasistencia=recordantiguo.completoasistencia,
                                                                  fechainicio=recordantiguo.fechainicio,
                                                                  fechafin=recordantiguo.fechafin,
                                                                  suficiencia=recordantiguo.suficiencia,
                                                                  asignaturamallahistorico_id=recordantiguo.asignaturamalla.id,
                                                                  reverso=False)
                                    recordnuevo.save()
                                    print(u"Crea nuevo record %s" % recordnuevo)


                                elif RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                    asignaturamalla=equivalencia.asignaturamallasalto):
                                    recordnuevo = RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                                 asignaturamalla=equivalencia.asignaturamallasalto)[0]
                                    recordnuevo.matriculas = recordantiguo.matriculas
                                    recordnuevo.asignaturamalla = equivalencia.asignaturamallasalto
                                    recordnuevo.asignatura = equivalencia.asignaturamallasalto.asignatura
                                    recordnuevo.asignaturaold = recordantiguo.asignaturamalla.asignatura
                                    recordnuevo.nota = recordantiguo.nota
                                    recordnuevo.asistencia = recordantiguo.asistencia
                                    recordnuevo.sinasistencia = recordantiguo.sinasistencia
                                    recordnuevo.fecha = recordantiguo.fecha
                                    recordnuevo.noaplica = recordantiguo.noaplica
                                    recordnuevo.aprobada = recordantiguo.aprobada
                                    recordnuevo.convalidacion = recordantiguo.convalidacion
                                    recordnuevo.pendiente = recordantiguo.pendiente
                                    recordnuevo.creditos = equivalencia.asignaturamallasalto.creditos
                                    recordnuevo.horas = equivalencia.asignaturamallasalto.horas
                                    recordnuevo.valida = recordantiguo.valida
                                    recordnuevo.validapromedio = recordantiguo.validapromedio
                                    recordnuevo.observaciones = observaciones
                                    recordnuevo.homologada = homologada
                                    recordnuevo.materiaregular = recordantiguo.materiaregular
                                    recordnuevo.materiacurso = None
                                    recordnuevo.completonota = recordantiguo.completonota
                                    recordnuevo.completoasistencia = recordantiguo.completoasistencia
                                    recordnuevo.fechainicio = recordantiguo.fechainicio
                                    recordnuevo.fechafin = recordantiguo.fechafin
                                    recordnuevo.suficiencia = recordantiguo.suficiencia
                                    recordnuevo.asignaturamallahistorico = recordantiguo.asignaturamalla
                                    recordnuevo.reverso = False
                                    recordnuevo.save()

                                if recordnuevo:
                                    historicos = HistoricoRecordAcademico.objects.filter(status=True,
                                                                                         recordacademico=recordantiguo).update(recordacademico=recordnuevo,
                                                                                                                               creditos=recordnuevo.creditos,
                                                                                                                               horas=recordnuevo.horas,
                                                                                                                               homologada=recordnuevo.homologada)
                                    respaldo = RespaldoRecordAcademico.objects.filter(status=True,recordacademicooriginal=recordantiguo)

                                    if equivalencia.asignaturamallasalto_id in [10850,10853,10854]:
                                        if not practicaspp:
                                            if recordnuevo.aprobada:
                                                profesor = None
                                                if recordnuevo.materiaregular:
                                                    profesor = recordnuevo.materiaregular.profesor_principal()
                                                elif recordnuevo.materiacurso:
                                                    profesor = recordnuevo.materiaregular.profesor()
                                                if equivalencia.asignaturamallasalto_id == 10850:
                                                    itinerariosanteriores = ItinerariosMalla.objects.filter(status=True, malla_id=mallaantigua_id, nivel_id=8).order_by('id')[0]
                                                    itinerariosnuevos = ItinerariosMalla.objects.filter(status=True, malla_id=mallanueva_id, nivel_id=8).order_by('id')[0]

                                                    practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                                   inscripcion=inscripcion,
                                                                                                                   estadosolicitud__in=[
                                                                                                                       1, 2, 4,
                                                                                                                       5, 6],
                                                                                                                   itinerariomalla=itinerariosanteriores).exists()
                                                    practicarechazada = False
                                                    if not practica:
                                                        practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            estadosolicitud=3,
                                                            itinerariomalla=itinerariosanteriores).exists()

                                                    if not practica or practicarechazada:
                                                        if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                                   inscripcion=inscripcion,
                                                                                                                   actividad__itinerariomalla=itinerariosanteriores).exists():
                                                            nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                                 inscripcion=inscripcion,
                                                                                                                 fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                                 fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                                 numerohora=itinerariosnuevos.horas_practicas,
                                                                                                                 nivelmalla=itinerariosnuevos.nivel,
                                                                                                                 tiposolicitud=1,
                                                                                                                 estadosolicitud=2,
                                                                                                                 tipo=1,
                                                                                                                 itinerariomalla=itinerariosnuevos,
                                                                                                                 supervisor=profesor,
                                                                                                                 tutorunemi=profesor,
                                                                                                                 fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                                 tipoinstitucion=1,
                                                                                                                 sectoreconomico=6,
                                                                                                                 empresaempleadora_id=3,
                                                                                                                 culminada=True,
                                                                                                                 fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                                 lugarpractica_id=2,
                                                                                                                 observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                                 )
                                                            nuevapractica.save()

                                                    ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                                  itinerario=itinerariosanteriores).update(
                                                        itinerario=itinerariosnuevos)

                                                if equivalencia.asignaturamallasalto_id == 10854:
                                                    itinerariosanteriores = ItinerariosMalla.objects.filter(status=True, malla_id=mallaantigua_id, nivel_id=8).order_by('id')[1]
                                                    itinerariosnuevos = ItinerariosMalla.objects.filter(status=True, malla_id=mallanueva_id, nivel_id=8).order_by('id')[1]

                                                    practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                                   inscripcion=inscripcion,
                                                                                                                   estadosolicitud__in=[
                                                                                                                       1, 2, 4,
                                                                                                                       5, 6],
                                                                                                                   itinerariomalla=itinerariosanteriores).exists()
                                                    practicarechazada = False
                                                    if not practica:
                                                        practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            estadosolicitud=3,
                                                            itinerariomalla=itinerariosanteriores).exists()

                                                    if not practica or practicarechazada:
                                                        if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                                   inscripcion=inscripcion,
                                                                                                                   actividad__itinerariomalla=itinerariosanteriores).exists():
                                                            nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                                 inscripcion=inscripcion,
                                                                                                                 fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                                 fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                                 numerohora=itinerariosnuevos.horas_practicas,
                                                                                                                 nivelmalla=itinerariosnuevos.nivel,
                                                                                                                 tiposolicitud=1,
                                                                                                                 estadosolicitud=2,
                                                                                                                 tipo=1,
                                                                                                                 itinerariomalla=itinerariosnuevos,
                                                                                                                 supervisor=profesor,
                                                                                                                 tutorunemi=profesor,
                                                                                                                 fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                                 tipoinstitucion=1,
                                                                                                                 sectoreconomico=6,
                                                                                                                 empresaempleadora_id=3,
                                                                                                                 culminada=True,
                                                                                                                 fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                                 lugarpractica_id=2,
                                                                                                                 observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                                 )
                                                            nuevapractica.save()

                                                    ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                                  itinerario=itinerariosanteriores).update(
                                                        itinerario=itinerariosnuevos)

                                                if equivalencia.asignaturamallasalto_id == 10853:
                                                    itinerariosanteriores = ItinerariosMalla.objects.filter(status=True, malla_id=mallaantigua_id, nivel_id=8).order_by('id')[2]
                                                    itinerariosnuevos = ItinerariosMalla.objects.filter(status=True, malla_id=mallanueva_id, nivel_id=8).order_by('id')[2]

                                                    practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                                   inscripcion=inscripcion,
                                                                                                                   estadosolicitud__in=[
                                                                                                                       1, 2, 4,
                                                                                                                       5, 6],
                                                                                                                   itinerariomalla=itinerariosanteriores).exists()
                                                    practicarechazada = False
                                                    if not practica:
                                                        practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            estadosolicitud=3,
                                                            itinerariomalla=itinerariosanteriores).exists()

                                                    if not practica or practicarechazada:
                                                        if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                                   inscripcion=inscripcion,
                                                                                                                   actividad__itinerariomalla=itinerariosanteriores).exists():
                                                            nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                                 inscripcion=inscripcion,
                                                                                                                 fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                                 fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                                 numerohora=itinerariosnuevos.horas_practicas,
                                                                                                                 nivelmalla=itinerariosnuevos.nivel,
                                                                                                                 tiposolicitud=1,
                                                                                                                 estadosolicitud=2,
                                                                                                                 tipo=1,
                                                                                                                 itinerariomalla=itinerariosnuevos,
                                                                                                                 supervisor=profesor,
                                                                                                                 tutorunemi=profesor,
                                                                                                                 fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                                 tipoinstitucion=1,
                                                                                                                 sectoreconomico=6,
                                                                                                                 empresaempleadora_id=3,
                                                                                                                 culminada=True,
                                                                                                                 fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                                 lugarpractica_id=2,
                                                                                                                 observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                                 )
                                                            nuevapractica.save()

                                                    ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                                  itinerario=itinerariosanteriores).update(
                                                        itinerario=itinerariosnuevos)


                                    if equivalencia.asignaturamallasalto_id in [10859,10865]:
                                        if not horasvinculacion:
                                            if recordnuevo.aprobada:
                                                if equivalencia.asignaturamallasalto_id == 10859 and inscripcion.numero_horas_proyectos_vinculacion() < 80:
                                                    horasfalta = 80 - inscripcion.numero_horas_proyectos_vinculacion()
                                                    vinculacion = ParticipantesMatrices(status=True,
                                                                                        matrizevidencia_id=2,
                                                                                        proyecto_id=601,
                                                                                        inscripcion=inscripcion,
                                                                                        horas=horasfalta,
                                                                                        registrohorasdesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                        registrohorashasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                        estado=1
                                                                                        )
                                                    vinculacion.save()

                                                if equivalencia.asignaturamallasalto_id == 10865 and inscripcion.numero_horas_proyectos_vinculacion() < 160:
                                                    horasfalta = 160 - inscripcion.numero_horas_proyectos_vinculacion()
                                                    vinculacion = ParticipantesMatrices(status=True,
                                                                                        matrizevidencia_id=2,
                                                                                        proyecto_id=601,
                                                                                        inscripcion=inscripcion,
                                                                                        horas=horasfalta,
                                                                                        registrohorasdesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                        registrohorashasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                        estado=1
                                                                                        )
                                                    vinculacion.save()

                                    if not respaldo.exists():
                                        respaldorecord = RespaldoRecordAcademico(
                                            recordacademicooriginal=recordantiguo,
                                            recordacademiconuevo=recordnuevo
                                        )
                                        respaldorecord.save()
                                    else:
                                        respaldorecord = respaldo[0]
                                        respaldorecord.recordacademiconuevo = recordnuevo
                                        respaldorecord.save()
                                    print(u"Record actualizado %s" % recordnuevo)

                            else:
                                hojadestino.write(fila, 3, "NO ENCONTRO RECORD ANTIGUO %s" % equivalencia.asignaturamalla)
                                fila += 1

                        # if cont_asig_vinculacion_aprobadas != 0:
                        #     if cont_asig_vinculacion_aprobadas == 1:
                        #         horasfalta = 80
                        #     elif cont_asig_vinculacion_aprobadas == 2:
                        #         horasfalta = 160
                        #     vinculacion = ParticipantesMatrices(status=True,
                        #                                         matrizevidencia_id=2,
                        #                                         proyecto_id=601,
                        #                                         inscripcion=inscripcion,
                        #                                         horas=horasfalta,
                        #                                         registrohorasdesde=datetime.now().date(),
                        #                                         registrohorashasta=datetime.now().date(),
                        #                                         estado=1
                        #                                         )
                        #     vinculacion.save()

                        practicasppf = inscripcion.numero_horas_practicas_pre_profesionales()
                        hojadestino.write(fila, 3, practicasppf, fuentenormal)
                        horasvinculacionf = inscripcion.numero_horas_proyectos_vinculacion()
                        hojadestino.write(fila, 4, horasvinculacionf, fuentenormal)
                        fila += 1

                        time.sleep(1)

                    else:
                        sin_matricula.append(inscripcion)

            lin += 1

        libdestino.save(output_folder + libre_origen)
        print(output_folder + libre_origen)
        print("Proceso finalizado. . .")
        print(str(sin_matricula))

    except Exception as ex:
        transaction.set_rollback(True)
        print('error: %s' % ex)
        hojadestino.write(fila, 3, str(ex))
        fila += 1

def homologacion_ts_rezagados():

    #verificar que todos los estudiantes tengan la misma malla
    cadena = ''
    linea, excluidos, conexito = 0, 0, 0
    try:
        libre_origen = '/homologacion_ts_2_rezagados.xls'
        fuentecabecera = easyxf(
            'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
        fuentenormal = easyxf(
            'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')

        output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
        # liborigen = xlrd.open_workbook(output_folder + libre_origen)
        libdestino = xlwt.Workbook()
        hojadestino = libdestino.add_sheet('Sheet1')
        fil = 0
        columnas = [(u"CEDULA", 6000),
                   (u"APELLIDOS Y NOMBRES", 6000),
                   (u"OBSERVACIÓN", 6000),
                    (u"HORAS PRACTICAS", 6000),
                    (u"HORAS VINCULACION", 6000),
                    (u"OBSERVACIÓN", 6000)
        ]
        for col_num in range(len(columnas)):
            hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
            hojadestino.col(col_num).width = columnas[col_num][1]
        fila = 1
        titulacion = 0

        lin = 0
        #miarchivo = openpyxl.load_workbook("CEDULA_COMUNICACION_final.xlsx")
        miarchivo = openpyxl.load_workbook("inscripcion_ts.xlsx")
        #miarchivo = openpyxl.load_workbook("prueba_2.xlsx")

        ws = miarchivo.get_sheet_by_name("prueba")
        worksheet = ws
        c = 0
        cont = 0
        periodo_id=224
        carrera_id=130
        mallaantigua_id=206
        mallanueva_id=485
        sin_matricula = []

        for row in worksheet.iter_rows(min_row=0):
            if lin >= 0:
                currentValues, cadena = [], ''
                for cell in row:
                    cadena += str(cell.value) + ' '
                    currentValues.append(str(cell.value))
                inscripcion = int(currentValues[0])

                if not inscripcion:
                    break

                matricula = Matricula.objects.filter(status=True, inscripcion__carrera__id=carrera_id, inscripcion_id=inscripcion).first()
                if matricula:
                    cont += 1
                    matricula.pasoayuda = True
                    matricula.save()
                    print(u"%s - %s - %s" % (matricula, cont, inscripcion))
                    inscripcion = matricula.inscripcion
                    hojadestino.write(fila, 0, matricula.inscripcion.persona.identificacion(), fuentenormal)
                    hojadestino.write(fila, 1, matricula.inscripcion.persona.nombre_completo(), fuentenormal)
                    hojadestino.write(fila, 2, matricula.nivel.periodo.nombre, fuentenormal)

                    practicaspp= haber_cumplido_horas_creditos_practicas_preprofesionales(inscripcion.id)
                    horasvinculacion = haber_cumplido_horas_creditos_vinculacion(inscripcion.id)

                    if InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallaantigua_id).exists():
                        imantigua = InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallaantigua_id)[0]
                        imantigua.status = False
                        imantigua.save()
                        print(u"Desactiva antigua inscripcion -----------------------------")

                    if not InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallanueva_id).exists():
                        imnueva = InscripcionMalla(inscripcion=inscripcion, malla_id=mallanueva_id)
                        imnueva.save()
                        print(u"Crea nueva inscripcion -----------------------------")

                    equivalencias = TablaEquivalenciaAsignaturas.objects.filter(status=True, asignaturamalla__malla_id=mallaantigua_id).order_by('asignaturamallasalto__nivelmalla__orden')
                    for equivalencia in equivalencias:
                        print(u"nueva - %s" % equivalencia.asignaturamallasalto)
                        recordantiguo = inscripcion.recordacademico_set.filter(status=True,asignaturamalla_id=equivalencia.asignaturamalla_id).first()

                        if recordantiguo:
                            print(u"anterior - %s" % equivalencia.asignaturamalla)
                            print(u"Record antiguo: %s" % recordantiguo)
                            recordnuevo = None
                            recordantiguo.status = False
                            recordantiguo.save(update_asignaturamalla=False)

                            if equivalencia.asignaturamallasalto_id in [10724, 10730, 10777, 10787]:
                                observaciones = recordantiguo.observaciones + " Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15"
                                homologada = True



                            else:
                                observaciones = recordantiguo.observaciones + " Migración con base a RESOLUCIÓN CGA-SO-9-2023-NO15 Y RESOLUCIÓN OCS-SO-18-2023-NO14"
                                homologada = False
                            if not RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                  asignaturamalla=equivalencia.asignaturamallasalto).exists():



                                recordnuevo = RecordAcademico(inscripcion=inscripcion,
                                                              matriculas=recordantiguo.matriculas,
                                                              asignaturamalla=equivalencia.asignaturamallasalto,
                                                              asignatura=equivalencia.asignaturamallasalto.asignatura,
                                                              asignaturaold_id=recordantiguo.asignatura.id,
                                                              nota=recordantiguo.nota,
                                                              asistencia=recordantiguo.asistencia,
                                                              sinasistencia=recordantiguo.sinasistencia,
                                                              fecha=recordantiguo.fecha,
                                                              noaplica=recordantiguo.noaplica,
                                                              aprobada=recordantiguo.aprobada,
                                                              convalidacion=recordantiguo.convalidacion,
                                                              pendiente=recordantiguo.pendiente,
                                                              creditos=equivalencia.asignaturamallasalto.creditos,
                                                              horas=equivalencia.asignaturamallasalto.horas,
                                                              valida=recordantiguo.valida,
                                                              validapromedio=recordantiguo.validapromedio,
                                                              observaciones=observaciones,
                                                              homologada=homologada,
                                                              materiaregular=recordantiguo.materiaregular,
                                                              materiacurso=None,
                                                              completonota=recordantiguo.completonota,
                                                              completoasistencia=recordantiguo.completoasistencia,
                                                              fechainicio=recordantiguo.fechainicio,
                                                              fechafin=recordantiguo.fechafin,
                                                              suficiencia=recordantiguo.suficiencia,
                                                              asignaturamallahistorico_id=recordantiguo.asignaturamalla.id,
                                                              reverso=False)
                                recordnuevo.save()
                                print(u"Crea nuevo record %s" % recordnuevo)


                            elif RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                asignaturamalla=equivalencia.asignaturamallasalto):
                                recordnuevo = RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                             asignaturamalla=equivalencia.asignaturamallasalto)[0]
                                recordnuevo.matriculas = recordantiguo.matriculas
                                recordnuevo.asignaturamalla = equivalencia.asignaturamallasalto
                                recordnuevo.asignatura = equivalencia.asignaturamallasalto.asignatura
                                recordnuevo.asignaturaold = recordantiguo.asignaturamalla.asignatura
                                recordnuevo.nota = recordantiguo.nota
                                recordnuevo.asistencia = recordantiguo.asistencia
                                recordnuevo.sinasistencia = recordantiguo.sinasistencia
                                recordnuevo.fecha = recordantiguo.fecha
                                recordnuevo.noaplica = recordantiguo.noaplica
                                recordnuevo.aprobada = recordantiguo.aprobada
                                recordnuevo.convalidacion = recordantiguo.convalidacion
                                recordnuevo.pendiente = recordantiguo.pendiente
                                recordnuevo.creditos = equivalencia.asignaturamallasalto.creditos
                                recordnuevo.horas = equivalencia.asignaturamallasalto.horas
                                recordnuevo.valida = recordantiguo.valida
                                recordnuevo.validapromedio = recordantiguo.validapromedio
                                recordnuevo.observaciones = observaciones
                                recordnuevo.homologada = homologada
                                recordnuevo.materiaregular = recordantiguo.materiaregular
                                recordnuevo.materiacurso = None
                                recordnuevo.completonota = recordantiguo.completonota
                                recordnuevo.completoasistencia = recordantiguo.completoasistencia
                                recordnuevo.fechainicio = recordantiguo.fechainicio
                                recordnuevo.fechafin = recordantiguo.fechafin
                                recordnuevo.suficiencia = recordantiguo.suficiencia
                                recordnuevo.asignaturamallahistorico = recordantiguo.asignaturamalla
                                recordnuevo.reverso = False
                                recordnuevo.save()

                            if recordnuevo:
                                historicos = HistoricoRecordAcademico.objects.filter(status=True,
                                                                                     recordacademico=recordantiguo).update(recordacademico=recordnuevo,
                                                                                                                           creditos=recordnuevo.creditos,
                                                                                                                           horas=recordnuevo.horas,
                                                                                                                           homologada=recordnuevo.homologada)
                                respaldo = RespaldoRecordAcademico.objects.filter(status=True,recordacademicooriginal=recordantiguo)

                                if equivalencia.asignaturamallasalto_id in [10777, 10787]:
                                    if not practicaspp:
                                        if recordnuevo.aprobada:
                                            profesor = None
                                            if recordnuevo.materiaregular:
                                                profesor = recordnuevo.materiaregular.profesor_principal()
                                            elif recordnuevo.materiacurso:
                                                profesor = recordnuevo.materiaregular.profesor()
                                            if equivalencia.asignaturamallasalto_id == 10777:
                                                itinerariooctavo = ItinerariosMalla.objects.get(status=True, malla_id=mallaantigua_id, nivel_id=8)
                                                ioctavonuevo = ItinerariosMalla.objects.get(status=True, malla_id=mallanueva_id, nivel_id=8)


                                                practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                               inscripcion=inscripcion,
                                                                                                               estadosolicitud__in=[
                                                                                                                   1, 2, 4,
                                                                                                                   5, 6],
                                                                                                               itinerariomalla=itinerariooctavo).exists()
                                                practicarechazada = False
                                                if not practica:
                                                    practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        estadosolicitud=3,
                                                        itinerariomalla=itinerariooctavo).exists()


                                                if not practica or practicarechazada:
                                                    if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                               inscripcion=inscripcion,
                                                                                                               actividad__itinerariomalla=itinerariooctavo).exists():
                                                        nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                             inscripcion=inscripcion,
                                                                                                             fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             numerohora=ioctavonuevo.horas_practicas,
                                                                                                             nivelmalla=ioctavonuevo.nivel,
                                                                                                             tiposolicitud=1,
                                                                                                             estadosolicitud=2,
                                                                                                             tipo=1,
                                                                                                             itinerariomalla=ioctavonuevo,
                                                                                                             supervisor=profesor,
                                                                                                             tutorunemi=profesor,
                                                                                                             fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             tipoinstitucion=1,
                                                                                                             sectoreconomico=6,
                                                                                                             empresaempleadora_id=3,
                                                                                                             culminada=True,
                                                                                                             fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             lugarpractica_id=2,
                                                                                                             observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                             )
                                                        nuevapractica.save()

                                                ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                              itinerario=itinerariooctavo).update(
                                                    itinerario=ioctavonuevo)

                                            if equivalencia.asignaturamallasalto_id == 10787:
                                                itinerarionoveno = ItinerariosMalla.objects.get(status=True,
                                                                                               malla_id=mallaantigua_id,
                                                                                               nivel_id=9)
                                                inovenonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                            malla_id=mallanueva_id,
                                                                                            nivel_id=9)

                                                practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                               inscripcion=inscripcion,
                                                                                                               estadosolicitud__in=[
                                                                                                                   1, 2, 4,
                                                                                                                   5, 6],
                                                                                                               itinerariomalla=itinerarionoveno).exists()
                                                practicarechazada = False
                                                if not practica:
                                                    practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        estadosolicitud=3,
                                                        itinerariomalla=itinerarionoveno).exists()

                                                if not practica or practicarechazada:
                                                    if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                               inscripcion=inscripcion,
                                                                                                               actividad__itinerariomalla=itinerarionoveno).exists():
                                                        nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                             inscripcion=inscripcion,
                                                                                                             fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             numerohora=inovenonuevo.horas_practicas,
                                                                                                             nivelmalla=inovenonuevo.nivel,
                                                                                                             tiposolicitud=1,
                                                                                                             estadosolicitud=2,
                                                                                                             tipo=1,
                                                                                                             itinerariomalla=inovenonuevo,
                                                                                                             supervisor=profesor,
                                                                                                             tutorunemi=profesor,
                                                                                                             fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             tipoinstitucion=1,
                                                                                                             sectoreconomico=6,
                                                                                                             empresaempleadora_id=3,
                                                                                                             culminada=True,
                                                                                                             fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                             lugarpractica_id=2,
                                                                                                             observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                             )
                                                        nuevapractica.save()

                                                ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                              itinerario=itinerarionoveno).update(
                                                    itinerario=inovenonuevo)


                                if equivalencia.asignaturamallasalto_id in [10724, 10730]:
                                    if not horasvinculacion:
                                        if recordnuevo.aprobada:
                                            if equivalencia.asignaturamallasalto_id == 10724 and inscripcion.numero_horas_proyectos_vinculacion() < 80:
                                                horasfalta = 80 - inscripcion.numero_horas_proyectos_vinculacion()
                                                vinculacion = ParticipantesMatrices(status=True,
                                                                                    matrizevidencia_id=2,
                                                                                    proyecto_id=601,
                                                                                    inscripcion=inscripcion,
                                                                                    horas = horasfalta,
                                                                                    registrohorasdesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                    registrohorashasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                    estado=1
                                                )
                                                vinculacion.save()

                                            if equivalencia.asignaturamallasalto_id == 10730 and inscripcion.numero_horas_proyectos_vinculacion() < 160:
                                                horasfalta = 160 - inscripcion.numero_horas_proyectos_vinculacion()
                                                vinculacion = ParticipantesMatrices(status=True,
                                                                                    matrizevidencia_id=2,
                                                                                    proyecto_id=601,
                                                                                    inscripcion=inscripcion,
                                                                                    horas = horasfalta,
                                                                                    registrohorasdesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                    registrohorashasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                    estado=1
                                                )
                                                vinculacion.save()


                                if not respaldo.exists():
                                    respaldorecord = RespaldoRecordAcademico(
                                        recordacademicooriginal=recordantiguo,
                                        recordacademiconuevo=recordnuevo
                                    )
                                    respaldorecord.save()
                                else:
                                    respaldorecord = respaldo[0]
                                    respaldorecord.recordacademiconuevo = recordnuevo
                                    respaldorecord.save()
                                print(u"Record actualizado %s" % recordnuevo)


                        else:
                            hojadestino.write(fila, 3, "NO ENCONTRO RECORD ANTIGUO %s" % equivalencia.asignaturamalla)
                            fila += 1

                    practicasppf = inscripcion.numero_horas_practicas_pre_profesionales()
                    hojadestino.write(fila, 3, practicasppf, fuentenormal)
                    horasvinculacionf = inscripcion.numero_horas_proyectos_vinculacion()
                    hojadestino.write(fila, 4, horasvinculacionf, fuentenormal)

                    fila += 1

                    time.sleep(1)
                else:
                    sin_matricula.append(inscripcion)

            lin += 1

        libdestino.save(output_folder + libre_origen)
        print(output_folder + libre_origen)
        print("Proceso finalizado. . .")
        print(str(sin_matricula))


    except Exception as ex:
        transaction.set_rollback(True)
        print('error: %s' % ex)
        hojadestino.write(fila, 3, str(ex))
        fila += 1

def homologacion_tics():
    #verificar que todos los estudiantes tengan la misma malla
    cadena = ''
    linea, excluidos, conexito = 0, 0, 0
    try:
        libre_origen = '/homologacion_tics.xls'
        fuentecabecera = easyxf(
            'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
        fuentenormal = easyxf(
            'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')

        output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
        # liborigen = xlrd.open_workbook(output_folder + libre_origen)
        libdestino = xlwt.Workbook()
        hojadestino = libdestino.add_sheet('Sheet1')
        fil = 0
        columnas = [(u"CEDULA", 6000),
                    (u"APELLIDOS Y NOMBRES", 6000),
                    (u"OBSERVACIÓN", 6000),
                    (u"HORAS PRACTICAS", 6000),
                    (u"HORAS VINCULACION", 6000),
                    (u"OBSERVACIÓN", 6000)
                    ]
        for col_num in range(len(columnas)):
            hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
            hojadestino.col(col_num).width = columnas[col_num][1]
        fila = 1
        titulacion = 0

        lin = 0
        #miarchivo = openpyxl.load_workbook("CEDULA_COMUNICACION_final.xlsx")
        miarchivo = openpyxl.load_workbook("tecnologialinea.xlsx")
        #miarchivo = openpyxl.load_workbook("prueba_2.xlsx")

        ws = miarchivo.get_sheet_by_name("primero")
        worksheet = ws
        c = 0
        cont = 0
        periodo_id=224
        carrera_id=133
        mallaantigua_id=202
        mallanueva_id=478

        for row in worksheet.iter_rows(min_row=0):
            if lin >= 0:
                currentValues, cadena = [], ''
                for cell in row:
                    cadena += str(cell.value) + ' '
                    currentValues.append(str(cell.value))
                identificacion = currentValues[0]

                if not identificacion:
                    break

                with transaction.atomic():
                    matricula = Matricula.objects.filter(status=True, inscripcion__carrera__id=carrera_id, inscripcion__persona__cedula=identificacion).first()
                    cont += 1
                    matricula.pasoayuda = True
                    matricula.save()
                    print(u"%s - %s" % (matricula, cont))
                    inscripcion = matricula.inscripcion
                    hojadestino.write(fila, 0, matricula.inscripcion.persona.identificacion(), fuentenormal)
                    hojadestino.write(fila, 1, matricula.inscripcion.persona.nombre_completo(), fuentenormal)
                    hojadestino.write(fila, 2, matricula.nivel.periodo.nombre, fuentenormal)

                    practicaspp = haber_cumplido_horas_creditos_practicas_preprofesionales(inscripcion.id)
                    horasvinculacion = haber_cumplido_horas_creditos_vinculacion(inscripcion.id)

                    if InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallaantigua_id).exists():
                        imantigua = InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallaantigua_id)[0]
                        imantigua.status = False
                        imantigua.save()
                        print(u"Desactiva antigua inscripcion -----------------------------")

                    if not InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallanueva_id).exists():
                        imnueva = InscripcionMalla(inscripcion=inscripcion, malla_id=mallanueva_id)
                        imnueva.save()
                        print(u"Crea nueva inscripcion -----------------------------")

                    equivalencias = TablaEquivalenciaAsignaturas.objects.filter(status=True, asignaturamalla__malla_id=mallaantigua_id).order_by('asignaturamallasalto__nivelmalla__orden')
                    cont_asig_vinculacion_aprobadas = 0
                    horasfalta = 0
                    fechainicioitinerario = None
                    fechafinitinerario = None
                    for equivalencia in equivalencias:
                        print(u"nueva - %s" % equivalencia.asignaturamallasalto)
                        recordantiguo = inscripcion.recordacademico_set.filter(status=True,asignaturamalla_id=equivalencia.asignaturamalla_id).first()

                        if recordantiguo:
                            print(u"anterior - %s" % equivalencia.asignaturamalla)
                            print(u"Record antiguo: %s" % recordantiguo)
                            recordnuevo = None
                            recordantiguo.status = False
                            recordantiguo.save(update_asignaturamalla=False)

                            if equivalencia.asignaturamallasalto_id in [10587,10618,10612]:
                                observaciones = recordantiguo.observaciones + " Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15"
                                homologada = True


                            else:
                                observaciones = recordantiguo.observaciones + " Migración con base a RESOLUCIÓN CGA-SO-9-2023-NO15 Y RESOLUCIÓN OCS-SO-18-2023-NO14"
                                homologada = False
                            if not RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                  asignaturamalla=equivalencia.asignaturamallasalto).exists():



                                recordnuevo = RecordAcademico(inscripcion=inscripcion,
                                                              matriculas=recordantiguo.matriculas,
                                                              asignaturamalla=equivalencia.asignaturamallasalto,
                                                              asignatura=equivalencia.asignaturamallasalto.asignatura,
                                                              asignaturaold_id=recordantiguo.asignatura.id,
                                                              nota=recordantiguo.nota,
                                                              asistencia=recordantiguo.asistencia,
                                                              sinasistencia=recordantiguo.sinasistencia,
                                                              fecha=recordantiguo.fecha,
                                                              noaplica=recordantiguo.noaplica,
                                                              aprobada=recordantiguo.aprobada,
                                                              convalidacion=recordantiguo.convalidacion,
                                                              pendiente=recordantiguo.pendiente,
                                                              creditos=equivalencia.asignaturamallasalto.creditos,
                                                              horas=equivalencia.asignaturamallasalto.horas,
                                                              valida=recordantiguo.valida,
                                                              validapromedio=recordantiguo.validapromedio,
                                                              observaciones=observaciones,
                                                              homologada=homologada,
                                                              materiaregular=recordantiguo.materiaregular,
                                                              materiacurso=None,
                                                              completonota=recordantiguo.completonota,
                                                              completoasistencia=recordantiguo.completoasistencia,
                                                              fechainicio=recordantiguo.fechainicio,
                                                              fechafin=recordantiguo.fechafin,
                                                              suficiencia=recordantiguo.suficiencia,
                                                              asignaturamallahistorico_id=recordantiguo.asignaturamalla.id,
                                                              reverso=False)
                                recordnuevo.save()
                                print(u"Crea nuevo record %s" % recordnuevo)


                            elif RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                asignaturamalla=equivalencia.asignaturamallasalto):
                                recordnuevo = RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                             asignaturamalla=equivalencia.asignaturamallasalto)[0]
                                recordnuevo.matriculas = recordantiguo.matriculas
                                recordnuevo.asignaturamalla = equivalencia.asignaturamallasalto
                                recordnuevo.asignatura = equivalencia.asignaturamallasalto.asignatura
                                recordnuevo.asignaturaold = recordantiguo.asignaturamalla.asignatura
                                recordnuevo.nota = recordantiguo.nota
                                recordnuevo.asistencia = recordantiguo.asistencia
                                recordnuevo.sinasistencia = recordantiguo.sinasistencia
                                recordnuevo.fecha = recordantiguo.fecha
                                recordnuevo.noaplica = recordantiguo.noaplica
                                recordnuevo.aprobada = recordantiguo.aprobada
                                recordnuevo.convalidacion = recordantiguo.convalidacion
                                recordnuevo.pendiente = recordantiguo.pendiente
                                recordnuevo.creditos = equivalencia.asignaturamallasalto.creditos
                                recordnuevo.horas = equivalencia.asignaturamallasalto.horas
                                recordnuevo.valida = recordantiguo.valida
                                recordnuevo.validapromedio = recordantiguo.validapromedio
                                recordnuevo.observaciones = observaciones
                                recordnuevo.homologada = homologada
                                recordnuevo.materiaregular = recordantiguo.materiaregular
                                recordnuevo.materiacurso = None
                                recordnuevo.completonota = recordantiguo.completonota
                                recordnuevo.completoasistencia = recordantiguo.completoasistencia
                                recordnuevo.fechainicio = recordantiguo.fechainicio
                                recordnuevo.fechafin = recordantiguo.fechafin
                                recordnuevo.suficiencia = recordantiguo.suficiencia
                                recordnuevo.asignaturamallahistorico = recordantiguo.asignaturamalla
                                recordnuevo.reverso = False
                                recordnuevo.save()

                            if recordnuevo:
                                historicos = HistoricoRecordAcademico.objects.filter(status=True,
                                                                                     recordacademico=recordantiguo).update(recordacademico=recordnuevo,
                                                                                                                           creditos=recordnuevo.creditos,
                                                                                                                           horas=recordnuevo.horas,
                                                                                                                           homologada=recordnuevo.homologada)
                                respaldo = RespaldoRecordAcademico.objects.filter(status=True,recordacademicooriginal=recordantiguo)


                                if not respaldo.exists():
                                    respaldorecord = RespaldoRecordAcademico(
                                        recordacademicooriginal=recordantiguo,
                                        recordacademiconuevo=recordnuevo
                                    )
                                    respaldorecord.save()
                                else:
                                    respaldorecord = respaldo[0]
                                    respaldorecord.recordacademiconuevo = recordnuevo
                                    respaldorecord.save()
                                print(u"Record actualizado %s" % recordnuevo)

                        else:
                            hojadestino.write(fila, 3, "NO ENCONTRO RECORD ANTIGUO %s" % equivalencia.asignaturamalla)
                            fila += 1

                    practicasppf = inscripcion.numero_horas_practicas_pre_profesionales()
                    hojadestino.write(fila, 3, practicasppf, fuentenormal)
                    horasvinculacionf = inscripcion.numero_horas_proyectos_vinculacion()
                    hojadestino.write(fila, 4, horasvinculacionf, fuentenormal)
                    fila += 1

                    time.sleep(1)

            lin += 1

        libdestino.save(output_folder + libre_origen)
        print(output_folder + libre_origen)
        print("Proceso finalizado. . .")

    except Exception as ex:
        transaction.set_rollback(True)
        print('error: %s' % ex)
        hojadestino.write(fila, 3, str(ex))
        fila += 1


def homologacion_tics_2():
    # verificar que todos los estudiantes tengan la misma malla
    cadena = ''
    linea, excluidos, conexito = 0, 0, 0
    try:
        libre_origen = '/homologacion_tics_2.xls'
        fuentecabecera = easyxf(
            'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
        fuentenormal = easyxf(
            'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')

        output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
        # liborigen = xlrd.open_workbook(output_folder + libre_origen)
        libdestino = xlwt.Workbook()
        hojadestino = libdestino.add_sheet('Sheet1')
        fil = 0
        columnas = [(u"CEDULA", 6000),
                    (u"APELLIDOS Y NOMBRES", 6000),
                    (u"OBSERVACIÓN", 6000),
                    (u"HORAS PRACTICAS", 6000),
                    (u"HORAS VINCULACION", 6000),
                    (u"OBSERVACIÓN", 6000)
                    ]
        for col_num in range(len(columnas)):
            hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
            hojadestino.col(col_num).width = columnas[col_num][1]
        fila = 1
        titulacion = 0

        lin = 0
        # miarchivo = openpyxl.load_workbook("CEDULA_COMUNICACION_final.xlsx")
        miarchivo = openpyxl.load_workbook("tecnologialinea.xlsx")
        # miarchivo = openpyxl.load_workbook("prueba_2.xlsx")

        ws = miarchivo.get_sheet_by_name("segundo")
        worksheet = ws
        c = 0
        cont = 0
        periodo_id = 224
        carrera_id = 133
        mallaantigua_id = 202
        mallanueva_id = 478

        for row in worksheet.iter_rows(min_row=0):
            if lin >= 0:
                currentValues, cadena = [], ''
                for cell in row:
                    cadena += str(cell.value) + ' '
                    currentValues.append(str(cell.value))
                identificacion = currentValues[0]

                if not identificacion:
                    break

                with transaction.atomic():
                    matricula = Matricula.objects.filter(status=True, inscripcion__carrera__id=carrera_id,
                                                         inscripcion__persona__cedula=identificacion).first()
                    cont += 1
                    matricula.pasoayuda = True
                    matricula.save()
                    print(u"%s - %s" % (matricula, cont))
                    inscripcion = matricula.inscripcion
                    hojadestino.write(fila, 0, matricula.inscripcion.persona.identificacion(), fuentenormal)
                    hojadestino.write(fila, 1, matricula.inscripcion.persona.nombre_completo(), fuentenormal)
                    hojadestino.write(fila, 2, matricula.nivel.periodo.nombre, fuentenormal)

                    practicaspp = haber_cumplido_horas_creditos_practicas_preprofesionales(inscripcion.id)
                    horasvinculacion = haber_cumplido_horas_creditos_vinculacion(inscripcion.id)

                    if InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                       malla_id=mallaantigua_id).exists():
                        imantigua = \
                        InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallaantigua_id)[
                            0]
                        imantigua.status = False
                        imantigua.save()
                        print(u"Desactiva antigua inscripcion -----------------------------")

                    if not InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                           malla_id=mallanueva_id).exists():
                        imnueva = InscripcionMalla(inscripcion=inscripcion, malla_id=mallanueva_id)
                        imnueva.save()
                        print(u"Crea nueva inscripcion -----------------------------")

                    equivalencias = TablaEquivalenciaAsignaturas.objects.filter(status=True,
                                                                                asignaturamalla__malla_id=mallaantigua_id).order_by(
                        'asignaturamallasalto__nivelmalla__orden')
                    cont_asig_vinculacion_aprobadas = 0
                    horasfalta = 0
                    fechainicioitinerario = None
                    fechafinitinerario = None
                    for equivalencia in equivalencias:
                        print(u"nueva - %s" % equivalencia.asignaturamallasalto)
                        recordantiguo = inscripcion.recordacademico_set.filter(status=True,
                                                                               asignaturamalla_id=equivalencia.asignaturamalla_id).first()

                        if recordantiguo:
                            print(u"anterior - %s" % equivalencia.asignaturamalla)
                            print(u"Record antiguo: %s" % recordantiguo)
                            recordnuevo = None
                            recordantiguo.status = False
                            recordantiguo.save(update_asignaturamalla=False)

                            if equivalencia.asignaturamallasalto_id in [10587, 10618, 10612]:
                                observaciones = recordantiguo.observaciones + " Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15"
                                homologada = True


                            else:
                                observaciones = recordantiguo.observaciones + " Migración con base a RESOLUCIÓN CGA-SO-9-2023-NO15 Y RESOLUCIÓN OCS-SO-18-2023-NO14"
                                homologada = False
                            if not RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                  asignaturamalla=equivalencia.asignaturamallasalto).exists():

                                recordnuevo = RecordAcademico(inscripcion=inscripcion,
                                                              matriculas=recordantiguo.matriculas,
                                                              asignaturamalla=equivalencia.asignaturamallasalto,
                                                              asignatura=equivalencia.asignaturamallasalto.asignatura,
                                                              asignaturaold_id=recordantiguo.asignatura.id,
                                                              nota=recordantiguo.nota,
                                                              asistencia=recordantiguo.asistencia,
                                                              sinasistencia=recordantiguo.sinasistencia,
                                                              fecha=recordantiguo.fecha,
                                                              noaplica=recordantiguo.noaplica,
                                                              aprobada=recordantiguo.aprobada,
                                                              convalidacion=recordantiguo.convalidacion,
                                                              pendiente=recordantiguo.pendiente,
                                                              creditos=equivalencia.asignaturamallasalto.creditos,
                                                              horas=equivalencia.asignaturamallasalto.horas,
                                                              valida=recordantiguo.valida,
                                                              validapromedio=recordantiguo.validapromedio,
                                                              observaciones=observaciones,
                                                              homologada=homologada,
                                                              materiaregular=recordantiguo.materiaregular,
                                                              materiacurso=None,
                                                              completonota=recordantiguo.completonota,
                                                              completoasistencia=recordantiguo.completoasistencia,
                                                              fechainicio=recordantiguo.fechainicio,
                                                              fechafin=recordantiguo.fechafin,
                                                              suficiencia=recordantiguo.suficiencia,
                                                              asignaturamallahistorico_id=recordantiguo.asignaturamalla.id,
                                                              reverso=False)
                                recordnuevo.save()
                                print(u"Crea nuevo record %s" % recordnuevo)


                            elif RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                asignaturamalla=equivalencia.asignaturamallasalto):
                                recordnuevo = RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                             asignaturamalla=equivalencia.asignaturamallasalto)[
                                    0]
                                recordnuevo.matriculas = recordantiguo.matriculas
                                recordnuevo.asignaturamalla = equivalencia.asignaturamallasalto
                                recordnuevo.asignatura = equivalencia.asignaturamallasalto.asignatura
                                recordnuevo.asignaturaold = recordantiguo.asignaturamalla.asignatura
                                recordnuevo.nota = recordantiguo.nota
                                recordnuevo.asistencia = recordantiguo.asistencia
                                recordnuevo.sinasistencia = recordantiguo.sinasistencia
                                recordnuevo.fecha = recordantiguo.fecha
                                recordnuevo.noaplica = recordantiguo.noaplica
                                recordnuevo.aprobada = recordantiguo.aprobada
                                recordnuevo.convalidacion = recordantiguo.convalidacion
                                recordnuevo.pendiente = recordantiguo.pendiente
                                recordnuevo.creditos = equivalencia.asignaturamallasalto.creditos
                                recordnuevo.horas = equivalencia.asignaturamallasalto.horas
                                recordnuevo.valida = recordantiguo.valida
                                recordnuevo.validapromedio = recordantiguo.validapromedio
                                recordnuevo.observaciones = observaciones
                                recordnuevo.homologada = homologada
                                recordnuevo.materiaregular = recordantiguo.materiaregular
                                recordnuevo.materiacurso = None
                                recordnuevo.completonota = recordantiguo.completonota
                                recordnuevo.completoasistencia = recordantiguo.completoasistencia
                                recordnuevo.fechainicio = recordantiguo.fechainicio
                                recordnuevo.fechafin = recordantiguo.fechafin
                                recordnuevo.suficiencia = recordantiguo.suficiencia
                                recordnuevo.asignaturamallahistorico = recordantiguo.asignaturamalla
                                recordnuevo.reverso = False
                                recordnuevo.save()

                            if recordnuevo:
                                historicos = HistoricoRecordAcademico.objects.filter(status=True,
                                                                                     recordacademico=recordantiguo).update(
                                    recordacademico=recordnuevo,
                                    creditos=recordnuevo.creditos,
                                    horas=recordnuevo.horas,
                                    homologada=recordnuevo.homologada)
                                respaldo = RespaldoRecordAcademico.objects.filter(status=True,
                                                                                  recordacademicooriginal=recordantiguo)

                                if not respaldo.exists():
                                    respaldorecord = RespaldoRecordAcademico(
                                        recordacademicooriginal=recordantiguo,
                                        recordacademiconuevo=recordnuevo
                                    )
                                    respaldorecord.save()
                                else:
                                    respaldorecord = respaldo[0]
                                    respaldorecord.recordacademiconuevo = recordnuevo
                                    respaldorecord.save()
                                print(u"Record actualizado %s" % recordnuevo)

                        else:
                            hojadestino.write(fila, 3, "NO ENCONTRO RECORD ANTIGUO %s" % equivalencia.asignaturamalla)
                            fila += 1

                    practicasppf = inscripcion.numero_horas_practicas_pre_profesionales()
                    hojadestino.write(fila, 3, practicasppf, fuentenormal)
                    horasvinculacionf = inscripcion.numero_horas_proyectos_vinculacion()
                    hojadestino.write(fila, 4, horasvinculacionf, fuentenormal)
                    fila += 1

                    time.sleep(1)

            lin += 1

        libdestino.save(output_folder + libre_origen)
        print(output_folder + libre_origen)
        print("Proceso finalizado. . .")

    except Exception as ex:
        transaction.set_rollback(True)
        print('error: %s' % ex)
        hojadestino.write(fila, 3, str(ex))
        fila += 1
def homologacion_tics_3():
    # verificar que todos los estudiantes tengan la misma malla
    cadena = ''
    linea, excluidos, conexito = 0, 0, 0
    try:
        libre_origen = '/homologacion_tics_3.xls'
        fuentecabecera = easyxf(
            'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
        fuentenormal = easyxf(
            'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')

        output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
        # liborigen = xlrd.open_workbook(output_folder + libre_origen)
        libdestino = xlwt.Workbook()
        hojadestino = libdestino.add_sheet('Sheet1')
        fil = 0
        columnas = [(u"CEDULA", 6000),
                    (u"APELLIDOS Y NOMBRES", 6000),
                    (u"OBSERVACIÓN", 6000),
                    (u"HORAS PRACTICAS", 6000),
                    (u"HORAS VINCULACION", 6000),
                    (u"OBSERVACIÓN", 6000)
                    ]
        for col_num in range(len(columnas)):
            hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
            hojadestino.col(col_num).width = columnas[col_num][1]
        fila = 1
        titulacion = 0

        lin = 0
        # miarchivo = openpyxl.load_workbook("CEDULA_COMUNICACION_final.xlsx")
        miarchivo = openpyxl.load_workbook("tecnologialinea.xlsx")
        # miarchivo = openpyxl.load_workbook("prueba_2.xlsx")

        ws = miarchivo.get_sheet_by_name("tercero")
        worksheet = ws
        c = 0
        cont = 0
        periodo_id = 224
        carrera_id = 133
        mallaantigua_id = 202
        mallanueva_id = 478

        for row in worksheet.iter_rows(min_row=0):
            if lin >= 0:
                currentValues, cadena = [], ''
                for cell in row:
                    cadena += str(cell.value) + ' '
                    currentValues.append(str(cell.value))
                identificacion = currentValues[0]

                if not identificacion:
                    break

                with transaction.atomic():
                    matricula = Matricula.objects.filter(status=True, inscripcion__carrera__id=carrera_id,
                                                         inscripcion__persona__cedula=identificacion).first()
                    cont += 1
                    matricula.pasoayuda = True
                    matricula.save()
                    print(u"%s - %s" % (matricula, cont))
                    inscripcion = matricula.inscripcion
                    hojadestino.write(fila, 0, matricula.inscripcion.persona.identificacion(), fuentenormal)
                    hojadestino.write(fila, 1, matricula.inscripcion.persona.nombre_completo(), fuentenormal)
                    hojadestino.write(fila, 2, matricula.nivel.periodo.nombre, fuentenormal)

                    practicaspp = haber_cumplido_horas_creditos_practicas_preprofesionales(inscripcion.id)
                    horasvinculacion = haber_cumplido_horas_creditos_vinculacion(inscripcion.id)

                    if InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                       malla_id=mallaantigua_id).exists():
                        imantigua = \
                        InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallaantigua_id)[
                            0]
                        imantigua.status = False
                        imantigua.save()
                        print(u"Desactiva antigua inscripcion -----------------------------")

                    if not InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                           malla_id=mallanueva_id).exists():
                        imnueva = InscripcionMalla(inscripcion=inscripcion, malla_id=mallanueva_id)
                        imnueva.save()
                        print(u"Crea nueva inscripcion -----------------------------")

                    equivalencias = TablaEquivalenciaAsignaturas.objects.filter(status=True,
                                                                                asignaturamalla__malla_id=mallaantigua_id).order_by(
                        'asignaturamallasalto__nivelmalla__orden')
                    cont_asig_vinculacion_aprobadas = 0
                    horasfalta = 0
                    fechainicioitinerario = None
                    fechafinitinerario = None
                    for equivalencia in equivalencias:
                        print(u"nueva - %s" % equivalencia.asignaturamallasalto)
                        recordantiguo = inscripcion.recordacademico_set.filter(status=True,
                                                                               asignaturamalla_id=equivalencia.asignaturamalla_id).first()

                        if recordantiguo:
                            print(u"anterior - %s" % equivalencia.asignaturamalla)
                            print(u"Record antiguo: %s" % recordantiguo)
                            recordnuevo = None
                            recordantiguo.status = False
                            recordantiguo.save(update_asignaturamalla=False)

                            if equivalencia.asignaturamallasalto_id in [10587, 10618, 10612]:
                                observaciones = recordantiguo.observaciones + " Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15"
                                homologada = True


                            else:
                                observaciones = recordantiguo.observaciones + " Migración con base a RESOLUCIÓN CGA-SO-9-2023-NO15 Y RESOLUCIÓN OCS-SO-18-2023-NO14"
                                homologada = False
                            if not RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                  asignaturamalla=equivalencia.asignaturamallasalto).exists():

                                recordnuevo = RecordAcademico(inscripcion=inscripcion,
                                                              matriculas=recordantiguo.matriculas,
                                                              asignaturamalla=equivalencia.asignaturamallasalto,
                                                              asignatura=equivalencia.asignaturamallasalto.asignatura,
                                                              asignaturaold_id=recordantiguo.asignatura.id,
                                                              nota=recordantiguo.nota,
                                                              asistencia=recordantiguo.asistencia,
                                                              sinasistencia=recordantiguo.sinasistencia,
                                                              fecha=recordantiguo.fecha,
                                                              noaplica=recordantiguo.noaplica,
                                                              aprobada=recordantiguo.aprobada,
                                                              convalidacion=recordantiguo.convalidacion,
                                                              pendiente=recordantiguo.pendiente,
                                                              creditos=equivalencia.asignaturamallasalto.creditos,
                                                              horas=equivalencia.asignaturamallasalto.horas,
                                                              valida=recordantiguo.valida,
                                                              validapromedio=recordantiguo.validapromedio,
                                                              observaciones=observaciones,
                                                              homologada=homologada,
                                                              materiaregular=recordantiguo.materiaregular,
                                                              materiacurso=None,
                                                              completonota=recordantiguo.completonota,
                                                              completoasistencia=recordantiguo.completoasistencia,
                                                              fechainicio=recordantiguo.fechainicio,
                                                              fechafin=recordantiguo.fechafin,
                                                              suficiencia=recordantiguo.suficiencia,
                                                              asignaturamallahistorico_id=recordantiguo.asignaturamalla.id,
                                                              reverso=False)
                                recordnuevo.save()
                                print(u"Crea nuevo record %s" % recordnuevo)


                            elif RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                asignaturamalla=equivalencia.asignaturamallasalto):
                                recordnuevo = RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                             asignaturamalla=equivalencia.asignaturamallasalto)[
                                    0]
                                recordnuevo.matriculas = recordantiguo.matriculas
                                recordnuevo.asignaturamalla = equivalencia.asignaturamallasalto
                                recordnuevo.asignatura = equivalencia.asignaturamallasalto.asignatura
                                recordnuevo.asignaturaold = recordantiguo.asignaturamalla.asignatura
                                recordnuevo.nota = recordantiguo.nota
                                recordnuevo.asistencia = recordantiguo.asistencia
                                recordnuevo.sinasistencia = recordantiguo.sinasistencia
                                recordnuevo.fecha = recordantiguo.fecha
                                recordnuevo.noaplica = recordantiguo.noaplica
                                recordnuevo.aprobada = recordantiguo.aprobada
                                recordnuevo.convalidacion = recordantiguo.convalidacion
                                recordnuevo.pendiente = recordantiguo.pendiente
                                recordnuevo.creditos = equivalencia.asignaturamallasalto.creditos
                                recordnuevo.horas = equivalencia.asignaturamallasalto.horas
                                recordnuevo.valida = recordantiguo.valida
                                recordnuevo.validapromedio = recordantiguo.validapromedio
                                recordnuevo.observaciones = observaciones
                                recordnuevo.homologada = homologada
                                recordnuevo.materiaregular = recordantiguo.materiaregular
                                recordnuevo.materiacurso = None
                                recordnuevo.completonota = recordantiguo.completonota
                                recordnuevo.completoasistencia = recordantiguo.completoasistencia
                                recordnuevo.fechainicio = recordantiguo.fechainicio
                                recordnuevo.fechafin = recordantiguo.fechafin
                                recordnuevo.suficiencia = recordantiguo.suficiencia
                                recordnuevo.asignaturamallahistorico = recordantiguo.asignaturamalla
                                recordnuevo.reverso = False
                                recordnuevo.save()

                            if recordnuevo:
                                historicos = HistoricoRecordAcademico.objects.filter(status=True,
                                                                                     recordacademico=recordantiguo).update(
                                    recordacademico=recordnuevo,
                                    creditos=recordnuevo.creditos,
                                    horas=recordnuevo.horas,
                                    homologada=recordnuevo.homologada)
                                respaldo = RespaldoRecordAcademico.objects.filter(status=True,
                                                                                  recordacademicooriginal=recordantiguo)

                                if not respaldo.exists():
                                    respaldorecord = RespaldoRecordAcademico(
                                        recordacademicooriginal=recordantiguo,
                                        recordacademiconuevo=recordnuevo
                                    )
                                    respaldorecord.save()
                                else:
                                    respaldorecord = respaldo[0]
                                    respaldorecord.recordacademiconuevo = recordnuevo
                                    respaldorecord.save()
                                print(u"Record actualizado %s" % recordnuevo)

                        else:
                            hojadestino.write(fila, 3, "NO ENCONTRO RECORD ANTIGUO %s" % equivalencia.asignaturamalla)
                            fila += 1

                    practicasppf = inscripcion.numero_horas_practicas_pre_profesionales()
                    hojadestino.write(fila, 3, practicasppf, fuentenormal)
                    horasvinculacionf = inscripcion.numero_horas_proyectos_vinculacion()
                    hojadestino.write(fila, 4, horasvinculacionf, fuentenormal)
                    fila += 1

                    time.sleep(1)

            lin += 1

        libdestino.save(output_folder + libre_origen)
        print(output_folder + libre_origen)
        print("Proceso finalizado. . .")

    except Exception as ex:
        transaction.set_rollback(True)
        print('error: %s' % ex)
        hojadestino.write(fila, 3, str(ex))
        fila += 1
def homologacion_tics_4():
    # verificar que todos los estudiantes tengan la misma malla
    cadena = ''
    linea, excluidos, conexito = 0, 0, 0
    try:
        libre_origen = '/homologacion_tics_4.xls'
        fuentecabecera = easyxf(
            'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
        fuentenormal = easyxf(
            'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')

        output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
        # liborigen = xlrd.open_workbook(output_folder + libre_origen)
        libdestino = xlwt.Workbook()
        hojadestino = libdestino.add_sheet('Sheet1')
        fil = 0
        columnas = [(u"CEDULA", 6000),
                    (u"APELLIDOS Y NOMBRES", 6000),
                    (u"OBSERVACIÓN", 6000),
                    (u"HORAS PRACTICAS", 6000),
                    (u"HORAS VINCULACION", 6000),
                    (u"OBSERVACIÓN", 6000)
                    ]
        for col_num in range(len(columnas)):
            hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
            hojadestino.col(col_num).width = columnas[col_num][1]
        fila = 1
        titulacion = 0

        lin = 0
        # miarchivo = openpyxl.load_workbook("CEDULA_COMUNICACION_final.xlsx")
        miarchivo = openpyxl.load_workbook("tecnologialinea.xlsx")
        # miarchivo = openpyxl.load_workbook("prueba_2.xlsx")

        ws = miarchivo.get_sheet_by_name("cuarto")
        worksheet = ws
        c = 0
        cont = 0
        periodo_id = 224
        carrera_id = 133
        mallaantigua_id = 202
        mallanueva_id = 478

        for row in worksheet.iter_rows(min_row=0):
            if lin >= 0:
                currentValues, cadena = [], ''
                for cell in row:
                    cadena += str(cell.value) + ' '
                    currentValues.append(str(cell.value))
                identificacion = currentValues[0]

                if not identificacion:
                    break

                with transaction.atomic():
                    matricula = Matricula.objects.filter(status=True, inscripcion__carrera__id=carrera_id,
                                                         inscripcion__persona__cedula=identificacion).first()
                    cont += 1
                    matricula.pasoayuda = True
                    matricula.save()
                    print(u"%s - %s" % (matricula, cont))
                    inscripcion = matricula.inscripcion
                    hojadestino.write(fila, 0, matricula.inscripcion.persona.identificacion(), fuentenormal)
                    hojadestino.write(fila, 1, matricula.inscripcion.persona.nombre_completo(), fuentenormal)
                    hojadestino.write(fila, 2, matricula.nivel.periodo.nombre, fuentenormal)

                    practicaspp = haber_cumplido_horas_creditos_practicas_preprofesionales(inscripcion.id)
                    horasvinculacion = haber_cumplido_horas_creditos_vinculacion(inscripcion.id)

                    if InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                       malla_id=mallaantigua_id).exists():
                        imantigua = \
                        InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallaantigua_id)[
                            0]
                        imantigua.status = False
                        imantigua.save()
                        print(u"Desactiva antigua inscripcion -----------------------------")

                    if not InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                           malla_id=mallanueva_id).exists():
                        imnueva = InscripcionMalla(inscripcion=inscripcion, malla_id=mallanueva_id)
                        imnueva.save()
                        print(u"Crea nueva inscripcion -----------------------------")

                    equivalencias = TablaEquivalenciaAsignaturas.objects.filter(status=True,
                                                                                asignaturamalla__malla_id=mallaantigua_id).order_by(
                        'asignaturamallasalto__nivelmalla__orden')
                    cont_asig_vinculacion_aprobadas = 0
                    horasfalta = 0
                    fechainicioitinerario = None
                    fechafinitinerario = None
                    for equivalencia in equivalencias:
                        print(u"nueva - %s" % equivalencia.asignaturamallasalto)
                        recordantiguo = inscripcion.recordacademico_set.filter(status=True,
                                                                               asignaturamalla_id=equivalencia.asignaturamalla_id).first()

                        if recordantiguo:
                            print(u"anterior - %s" % equivalencia.asignaturamalla)
                            print(u"Record antiguo: %s" % recordantiguo)
                            recordnuevo = None
                            recordantiguo.status = False
                            recordantiguo.save(update_asignaturamalla=False)

                            if equivalencia.asignaturamallasalto_id in [10587, 10618, 10612]:
                                observaciones = recordantiguo.observaciones + " Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15"
                                homologada = True


                            else:
                                observaciones = recordantiguo.observaciones + " Migración con base a RESOLUCIÓN CGA-SO-9-2023-NO15 Y RESOLUCIÓN OCS-SO-18-2023-NO14"
                                homologada = False
                            if not RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                  asignaturamalla=equivalencia.asignaturamallasalto).exists():

                                recordnuevo = RecordAcademico(inscripcion=inscripcion,
                                                              matriculas=recordantiguo.matriculas,
                                                              asignaturamalla=equivalencia.asignaturamallasalto,
                                                              asignatura=equivalencia.asignaturamallasalto.asignatura,
                                                              asignaturaold_id=recordantiguo.asignatura.id,
                                                              nota=recordantiguo.nota,
                                                              asistencia=recordantiguo.asistencia,
                                                              sinasistencia=recordantiguo.sinasistencia,
                                                              fecha=recordantiguo.fecha,
                                                              noaplica=recordantiguo.noaplica,
                                                              aprobada=recordantiguo.aprobada,
                                                              convalidacion=recordantiguo.convalidacion,
                                                              pendiente=recordantiguo.pendiente,
                                                              creditos=equivalencia.asignaturamallasalto.creditos,
                                                              horas=equivalencia.asignaturamallasalto.horas,
                                                              valida=recordantiguo.valida,
                                                              validapromedio=recordantiguo.validapromedio,
                                                              observaciones=observaciones,
                                                              homologada=homologada,
                                                              materiaregular=recordantiguo.materiaregular,
                                                              materiacurso=None,
                                                              completonota=recordantiguo.completonota,
                                                              completoasistencia=recordantiguo.completoasistencia,
                                                              fechainicio=recordantiguo.fechainicio,
                                                              fechafin=recordantiguo.fechafin,
                                                              suficiencia=recordantiguo.suficiencia,
                                                              asignaturamallahistorico_id=recordantiguo.asignaturamalla.id,
                                                              reverso=False)
                                recordnuevo.save()
                                print(u"Crea nuevo record %s" % recordnuevo)


                            elif RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                asignaturamalla=equivalencia.asignaturamallasalto):
                                recordnuevo = RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                             asignaturamalla=equivalencia.asignaturamallasalto)[
                                    0]
                                recordnuevo.matriculas = recordantiguo.matriculas
                                recordnuevo.asignaturamalla = equivalencia.asignaturamallasalto
                                recordnuevo.asignatura = equivalencia.asignaturamallasalto.asignatura
                                recordnuevo.asignaturaold = recordantiguo.asignaturamalla.asignatura
                                recordnuevo.nota = recordantiguo.nota
                                recordnuevo.asistencia = recordantiguo.asistencia
                                recordnuevo.sinasistencia = recordantiguo.sinasistencia
                                recordnuevo.fecha = recordantiguo.fecha
                                recordnuevo.noaplica = recordantiguo.noaplica
                                recordnuevo.aprobada = recordantiguo.aprobada
                                recordnuevo.convalidacion = recordantiguo.convalidacion
                                recordnuevo.pendiente = recordantiguo.pendiente
                                recordnuevo.creditos = equivalencia.asignaturamallasalto.creditos
                                recordnuevo.horas = equivalencia.asignaturamallasalto.horas
                                recordnuevo.valida = recordantiguo.valida
                                recordnuevo.validapromedio = recordantiguo.validapromedio
                                recordnuevo.observaciones = observaciones
                                recordnuevo.homologada = homologada
                                recordnuevo.materiaregular = recordantiguo.materiaregular
                                recordnuevo.materiacurso = None
                                recordnuevo.completonota = recordantiguo.completonota
                                recordnuevo.completoasistencia = recordantiguo.completoasistencia
                                recordnuevo.fechainicio = recordantiguo.fechainicio
                                recordnuevo.fechafin = recordantiguo.fechafin
                                recordnuevo.suficiencia = recordantiguo.suficiencia
                                recordnuevo.asignaturamallahistorico = recordantiguo.asignaturamalla
                                recordnuevo.reverso = False
                                recordnuevo.save()

                            if recordnuevo:
                                historicos = HistoricoRecordAcademico.objects.filter(status=True,
                                                                                     recordacademico=recordantiguo).update(
                                    recordacademico=recordnuevo,
                                    creditos=recordnuevo.creditos,
                                    horas=recordnuevo.horas,
                                    homologada=recordnuevo.homologada)
                                respaldo = RespaldoRecordAcademico.objects.filter(status=True,
                                                                                  recordacademicooriginal=recordantiguo)

                                if not respaldo.exists():
                                    respaldorecord = RespaldoRecordAcademico(
                                        recordacademicooriginal=recordantiguo,
                                        recordacademiconuevo=recordnuevo
                                    )
                                    respaldorecord.save()
                                else:
                                    respaldorecord = respaldo[0]
                                    respaldorecord.recordacademiconuevo = recordnuevo
                                    respaldorecord.save()
                                print(u"Record actualizado %s" % recordnuevo)

                        else:
                            hojadestino.write(fila, 3, "NO ENCONTRO RECORD ANTIGUO %s" % equivalencia.asignaturamalla)
                            fila += 1

                    practicasppf = inscripcion.numero_horas_practicas_pre_profesionales()
                    hojadestino.write(fila, 3, practicasppf, fuentenormal)
                    horasvinculacionf = inscripcion.numero_horas_proyectos_vinculacion()
                    hojadestino.write(fila, 4, horasvinculacionf, fuentenormal)
                    fila += 1

                    time.sleep(1)

            lin += 1

        libdestino.save(output_folder + libre_origen)
        print(output_folder + libre_origen)
        print("Proceso finalizado. . .")

    except Exception as ex:
        transaction.set_rollback(True)
        print('error: %s' % ex)
        hojadestino.write(fila, 3, str(ex))
        fila += 1
def homologacion_tics_5():
    # verificar que todos los estudiantes tengan la misma malla
    cadena = ''
    linea, excluidos, conexito = 0, 0, 0
    try:
        libre_origen = '/homologacion_tics_5.xls'
        fuentecabecera = easyxf(
            'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
        fuentenormal = easyxf(
            'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')

        output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
        # liborigen = xlrd.open_workbook(output_folder + libre_origen)
        libdestino = xlwt.Workbook()
        hojadestino = libdestino.add_sheet('Sheet1')
        fil = 0
        columnas = [(u"CEDULA", 6000),
                    (u"APELLIDOS Y NOMBRES", 6000),
                    (u"OBSERVACIÓN", 6000),
                    (u"HORAS PRACTICAS", 6000),
                    (u"HORAS VINCULACION", 6000),
                    (u"OBSERVACIÓN", 6000)
                    ]
        for col_num in range(len(columnas)):
            hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
            hojadestino.col(col_num).width = columnas[col_num][1]
        fila = 1
        titulacion = 0

        lin = 0
        # miarchivo = openpyxl.load_workbook("CEDULA_COMUNICACION_final.xlsx")
        miarchivo = openpyxl.load_workbook("tecnologialinea.xlsx")
        # miarchivo = openpyxl.load_workbook("prueba_2.xlsx")

        ws = miarchivo.get_sheet_by_name("quinto")
        worksheet = ws
        c = 0
        cont = 0
        periodo_id = 224
        carrera_id = 133
        mallaantigua_id = 202
        mallanueva_id = 478

        for row in worksheet.iter_rows(min_row=0):
            if lin >= 0:
                currentValues, cadena = [], ''
                for cell in row:
                    cadena += str(cell.value) + ' '
                    currentValues.append(str(cell.value))
                identificacion = currentValues[0]

                if not identificacion:
                    break

                with transaction.atomic():
                    matricula = Matricula.objects.filter(status=True, inscripcion__carrera__id=carrera_id,
                                                         inscripcion__persona__cedula=identificacion).first()
                    cont += 1
                    matricula.pasoayuda = True
                    matricula.save()
                    print(u"%s - %s" % (matricula, cont))
                    inscripcion = matricula.inscripcion
                    hojadestino.write(fila, 0, matricula.inscripcion.persona.identificacion(), fuentenormal)
                    hojadestino.write(fila, 1, matricula.inscripcion.persona.nombre_completo(), fuentenormal)
                    hojadestino.write(fila, 2, matricula.nivel.periodo.nombre, fuentenormal)

                    practicaspp = haber_cumplido_horas_creditos_practicas_preprofesionales(inscripcion.id)
                    horasvinculacion = haber_cumplido_horas_creditos_vinculacion(inscripcion.id)

                    if InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                       malla_id=mallaantigua_id).exists():
                        imantigua = \
                        InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallaantigua_id)[
                            0]
                        imantigua.status = False
                        imantigua.save()
                        print(u"Desactiva antigua inscripcion -----------------------------")

                    if not InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                           malla_id=mallanueva_id).exists():
                        imnueva = InscripcionMalla(inscripcion=inscripcion, malla_id=mallanueva_id)
                        imnueva.save()
                        print(u"Crea nueva inscripcion -----------------------------")

                    equivalencias = TablaEquivalenciaAsignaturas.objects.filter(status=True,
                                                                                asignaturamalla__malla_id=mallaantigua_id).order_by(
                        'asignaturamallasalto__nivelmalla__orden')
                    cont_asig_vinculacion_aprobadas = 0
                    horasfalta = 0
                    fechainicioitinerario = None
                    fechafinitinerario = None
                    for equivalencia in equivalencias:
                        print(u"nueva - %s" % equivalencia.asignaturamallasalto)
                        recordantiguo = inscripcion.recordacademico_set.filter(status=True,
                                                                               asignaturamalla_id=equivalencia.asignaturamalla_id).first()

                        if recordantiguo:
                            print(u"anterior - %s" % equivalencia.asignaturamalla)
                            print(u"Record antiguo: %s" % recordantiguo)
                            recordnuevo = None
                            recordantiguo.status = False
                            recordantiguo.save(update_asignaturamalla=False)

                            if equivalencia.asignaturamallasalto_id in [10587, 10618, 10612]:
                                observaciones = recordantiguo.observaciones + " Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15"
                                homologada = True


                            else:
                                observaciones = recordantiguo.observaciones + " Migración con base a RESOLUCIÓN CGA-SO-9-2023-NO15 Y RESOLUCIÓN OCS-SO-18-2023-NO14"
                                homologada = False
                            if not RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                  asignaturamalla=equivalencia.asignaturamallasalto).exists():

                                recordnuevo = RecordAcademico(inscripcion=inscripcion,
                                                              matriculas=recordantiguo.matriculas,
                                                              asignaturamalla=equivalencia.asignaturamallasalto,
                                                              asignatura=equivalencia.asignaturamallasalto.asignatura,
                                                              asignaturaold_id=recordantiguo.asignatura.id,
                                                              nota=recordantiguo.nota,
                                                              asistencia=recordantiguo.asistencia,
                                                              sinasistencia=recordantiguo.sinasistencia,
                                                              fecha=recordantiguo.fecha,
                                                              noaplica=recordantiguo.noaplica,
                                                              aprobada=recordantiguo.aprobada,
                                                              convalidacion=recordantiguo.convalidacion,
                                                              pendiente=recordantiguo.pendiente,
                                                              creditos=equivalencia.asignaturamallasalto.creditos,
                                                              horas=equivalencia.asignaturamallasalto.horas,
                                                              valida=recordantiguo.valida,
                                                              validapromedio=recordantiguo.validapromedio,
                                                              observaciones=observaciones,
                                                              homologada=homologada,
                                                              materiaregular=recordantiguo.materiaregular,
                                                              materiacurso=None,
                                                              completonota=recordantiguo.completonota,
                                                              completoasistencia=recordantiguo.completoasistencia,
                                                              fechainicio=recordantiguo.fechainicio,
                                                              fechafin=recordantiguo.fechafin,
                                                              suficiencia=recordantiguo.suficiencia,
                                                              asignaturamallahistorico_id=recordantiguo.asignaturamalla.id,
                                                              reverso=False)
                                recordnuevo.save()
                                print(u"Crea nuevo record %s" % recordnuevo)


                            elif RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                asignaturamalla=equivalencia.asignaturamallasalto):
                                recordnuevo = RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                             asignaturamalla=equivalencia.asignaturamallasalto)[
                                    0]
                                recordnuevo.matriculas = recordantiguo.matriculas
                                recordnuevo.asignaturamalla = equivalencia.asignaturamallasalto
                                recordnuevo.asignatura = equivalencia.asignaturamallasalto.asignatura
                                recordnuevo.asignaturaold = recordantiguo.asignaturamalla.asignatura
                                recordnuevo.nota = recordantiguo.nota
                                recordnuevo.asistencia = recordantiguo.asistencia
                                recordnuevo.sinasistencia = recordantiguo.sinasistencia
                                recordnuevo.fecha = recordantiguo.fecha
                                recordnuevo.noaplica = recordantiguo.noaplica
                                recordnuevo.aprobada = recordantiguo.aprobada
                                recordnuevo.convalidacion = recordantiguo.convalidacion
                                recordnuevo.pendiente = recordantiguo.pendiente
                                recordnuevo.creditos = equivalencia.asignaturamallasalto.creditos
                                recordnuevo.horas = equivalencia.asignaturamallasalto.horas
                                recordnuevo.valida = recordantiguo.valida
                                recordnuevo.validapromedio = recordantiguo.validapromedio
                                recordnuevo.observaciones = observaciones
                                recordnuevo.homologada = homologada
                                recordnuevo.materiaregular = recordantiguo.materiaregular
                                recordnuevo.materiacurso = None
                                recordnuevo.completonota = recordantiguo.completonota
                                recordnuevo.completoasistencia = recordantiguo.completoasistencia
                                recordnuevo.fechainicio = recordantiguo.fechainicio
                                recordnuevo.fechafin = recordantiguo.fechafin
                                recordnuevo.suficiencia = recordantiguo.suficiencia
                                recordnuevo.asignaturamallahistorico = recordantiguo.asignaturamalla
                                recordnuevo.reverso = False
                                recordnuevo.save()

                            if recordnuevo:
                                historicos = HistoricoRecordAcademico.objects.filter(status=True,
                                                                                     recordacademico=recordantiguo).update(
                                    recordacademico=recordnuevo,
                                    creditos=recordnuevo.creditos,
                                    horas=recordnuevo.horas,
                                    homologada=recordnuevo.homologada)
                                respaldo = RespaldoRecordAcademico.objects.filter(status=True,
                                                                                  recordacademicooriginal=recordantiguo)

                                if not respaldo.exists():
                                    respaldorecord = RespaldoRecordAcademico(
                                        recordacademicooriginal=recordantiguo,
                                        recordacademiconuevo=recordnuevo
                                    )
                                    respaldorecord.save()
                                else:
                                    respaldorecord = respaldo[0]
                                    respaldorecord.recordacademiconuevo = recordnuevo
                                    respaldorecord.save()
                                print(u"Record actualizado %s" % recordnuevo)

                        else:
                            hojadestino.write(fila, 3, "NO ENCONTRO RECORD ANTIGUO %s" % equivalencia.asignaturamalla)
                            fila += 1

                    practicasppf = inscripcion.numero_horas_practicas_pre_profesionales()
                    hojadestino.write(fila, 3, practicasppf, fuentenormal)
                    horasvinculacionf = inscripcion.numero_horas_proyectos_vinculacion()
                    hojadestino.write(fila, 4, horasvinculacionf, fuentenormal)
                    fila += 1

                    time.sleep(1)

            lin += 1

        libdestino.save(output_folder + libre_origen)
        print(output_folder + libre_origen)
        print("Proceso finalizado. . .")

    except Exception as ex:
        transaction.set_rollback(True)
        print('error: %s' % ex)
        hojadestino.write(fila, 3, str(ex))
        fila += 1
def homologacion_tics_6():
    # verificar que todos los estudiantes tengan la misma malla
    cadena = ''
    linea, excluidos, conexito = 0, 0, 0
    try:
        libre_origen = '/homologacion_tics_6.xls'
        fuentecabecera = easyxf(
            'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
        fuentenormal = easyxf(
            'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')

        output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
        # liborigen = xlrd.open_workbook(output_folder + libre_origen)
        libdestino = xlwt.Workbook()
        hojadestino = libdestino.add_sheet('Sheet1')
        fil = 0
        columnas = [(u"CEDULA", 6000),
                    (u"APELLIDOS Y NOMBRES", 6000),
                    (u"OBSERVACIÓN", 6000),
                    (u"HORAS PRACTICAS", 6000),
                    (u"HORAS VINCULACION", 6000),
                    (u"OBSERVACIÓN", 6000)
                    ]
        for col_num in range(len(columnas)):
            hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
            hojadestino.col(col_num).width = columnas[col_num][1]
        fila = 1
        titulacion = 0

        lin = 0
        # miarchivo = openpyxl.load_workbook("CEDULA_COMUNICACION_final.xlsx")
        miarchivo = openpyxl.load_workbook("tecnologialinea.xlsx")
        # miarchivo = openpyxl.load_workbook("prueba_2.xlsx")

        ws = miarchivo.get_sheet_by_name("sexto")
        worksheet = ws
        c = 0
        cont = 0
        periodo_id = 224
        carrera_id = 133
        mallaantigua_id = 202
        mallanueva_id = 478

        for row in worksheet.iter_rows(min_row=0):
            if lin >= 0:
                currentValues, cadena = [], ''
                for cell in row:
                    cadena += str(cell.value) + ' '
                    currentValues.append(str(cell.value))
                identificacion = currentValues[0]

                if not identificacion:
                    break

                with transaction.atomic():
                    matricula = Matricula.objects.filter(status=True, inscripcion__carrera__id=carrera_id,
                                                         inscripcion__persona__cedula=identificacion).first()
                    cont += 1
                    matricula.pasoayuda = True
                    matricula.save()
                    print(u"%s - %s" % (matricula, cont))
                    inscripcion = matricula.inscripcion
                    hojadestino.write(fila, 0, matricula.inscripcion.persona.identificacion(), fuentenormal)
                    hojadestino.write(fila, 1, matricula.inscripcion.persona.nombre_completo(), fuentenormal)
                    hojadestino.write(fila, 2, matricula.nivel.periodo.nombre, fuentenormal)

                    practicaspp = haber_cumplido_horas_creditos_practicas_preprofesionales(inscripcion.id)
                    horasvinculacion = haber_cumplido_horas_creditos_vinculacion(inscripcion.id)

                    if InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                       malla_id=mallaantigua_id).exists():
                        imantigua = \
                        InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallaantigua_id)[
                            0]
                        imantigua.status = False
                        imantigua.save()
                        print(u"Desactiva antigua inscripcion -----------------------------")

                    if not InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                           malla_id=mallanueva_id).exists():
                        imnueva = InscripcionMalla(inscripcion=inscripcion, malla_id=mallanueva_id)
                        imnueva.save()
                        print(u"Crea nueva inscripcion -----------------------------")

                    equivalencias = TablaEquivalenciaAsignaturas.objects.filter(status=True,
                                                                                asignaturamalla__malla_id=mallaantigua_id).order_by(
                        'asignaturamallasalto__nivelmalla__orden')
                    cont_asig_vinculacion_aprobadas = 0
                    horasfalta = 0
                    fechainicioitinerario = None
                    fechafinitinerario = None
                    for equivalencia in equivalencias:
                        print(u"nueva - %s" % equivalencia.asignaturamallasalto)
                        recordantiguo = inscripcion.recordacademico_set.filter(status=True,
                                                                               asignaturamalla_id=equivalencia.asignaturamalla_id).first()

                        if recordantiguo:
                            print(u"anterior - %s" % equivalencia.asignaturamalla)
                            print(u"Record antiguo: %s" % recordantiguo)
                            recordnuevo = None
                            recordantiguo.status = False
                            recordantiguo.save(update_asignaturamalla=False)

                            if equivalencia.asignaturamallasalto_id in [10587, 10618, 10612]:
                                observaciones = recordantiguo.observaciones + " Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15"
                                homologada = True


                            else:
                                observaciones = recordantiguo.observaciones + " Migración con base a RESOLUCIÓN CGA-SO-9-2023-NO15 Y RESOLUCIÓN OCS-SO-18-2023-NO14"
                                homologada = False
                            if not RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                  asignaturamalla=equivalencia.asignaturamallasalto).exists():

                                recordnuevo = RecordAcademico(inscripcion=inscripcion,
                                                              matriculas=recordantiguo.matriculas,
                                                              asignaturamalla=equivalencia.asignaturamallasalto,
                                                              asignatura=equivalencia.asignaturamallasalto.asignatura,
                                                              asignaturaold_id=recordantiguo.asignatura.id,
                                                              nota=recordantiguo.nota,
                                                              asistencia=recordantiguo.asistencia,
                                                              sinasistencia=recordantiguo.sinasistencia,
                                                              fecha=recordantiguo.fecha,
                                                              noaplica=recordantiguo.noaplica,
                                                              aprobada=recordantiguo.aprobada,
                                                              convalidacion=recordantiguo.convalidacion,
                                                              pendiente=recordantiguo.pendiente,
                                                              creditos=equivalencia.asignaturamallasalto.creditos,
                                                              horas=equivalencia.asignaturamallasalto.horas,
                                                              valida=recordantiguo.valida,
                                                              validapromedio=recordantiguo.validapromedio,
                                                              observaciones=observaciones,
                                                              homologada=homologada,
                                                              materiaregular=recordantiguo.materiaregular,
                                                              materiacurso=None,
                                                              completonota=recordantiguo.completonota,
                                                              completoasistencia=recordantiguo.completoasistencia,
                                                              fechainicio=recordantiguo.fechainicio,
                                                              fechafin=recordantiguo.fechafin,
                                                              suficiencia=recordantiguo.suficiencia,
                                                              asignaturamallahistorico_id=recordantiguo.asignaturamalla.id,
                                                              reverso=False)
                                recordnuevo.save()
                                print(u"Crea nuevo record %s" % recordnuevo)


                            elif RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                asignaturamalla=equivalencia.asignaturamallasalto):
                                recordnuevo = RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                             asignaturamalla=equivalencia.asignaturamallasalto)[
                                    0]
                                recordnuevo.matriculas = recordantiguo.matriculas
                                recordnuevo.asignaturamalla = equivalencia.asignaturamallasalto
                                recordnuevo.asignatura = equivalencia.asignaturamallasalto.asignatura
                                recordnuevo.asignaturaold = recordantiguo.asignaturamalla.asignatura
                                recordnuevo.nota = recordantiguo.nota
                                recordnuevo.asistencia = recordantiguo.asistencia
                                recordnuevo.sinasistencia = recordantiguo.sinasistencia
                                recordnuevo.fecha = recordantiguo.fecha
                                recordnuevo.noaplica = recordantiguo.noaplica
                                recordnuevo.aprobada = recordantiguo.aprobada
                                recordnuevo.convalidacion = recordantiguo.convalidacion
                                recordnuevo.pendiente = recordantiguo.pendiente
                                recordnuevo.creditos = equivalencia.asignaturamallasalto.creditos
                                recordnuevo.horas = equivalencia.asignaturamallasalto.horas
                                recordnuevo.valida = recordantiguo.valida
                                recordnuevo.validapromedio = recordantiguo.validapromedio
                                recordnuevo.observaciones = observaciones
                                recordnuevo.homologada = homologada
                                recordnuevo.materiaregular = recordantiguo.materiaregular
                                recordnuevo.materiacurso = None
                                recordnuevo.completonota = recordantiguo.completonota
                                recordnuevo.completoasistencia = recordantiguo.completoasistencia
                                recordnuevo.fechainicio = recordantiguo.fechainicio
                                recordnuevo.fechafin = recordantiguo.fechafin
                                recordnuevo.suficiencia = recordantiguo.suficiencia
                                recordnuevo.asignaturamallahistorico = recordantiguo.asignaturamalla
                                recordnuevo.reverso = False
                                recordnuevo.save()

                            if recordnuevo:
                                historicos = HistoricoRecordAcademico.objects.filter(status=True,
                                                                                     recordacademico=recordantiguo).update(
                                    recordacademico=recordnuevo,
                                    creditos=recordnuevo.creditos,
                                    horas=recordnuevo.horas,
                                    homologada=recordnuevo.homologada)
                                respaldo = RespaldoRecordAcademico.objects.filter(status=True,
                                                                                  recordacademicooriginal=recordantiguo)

                                if not respaldo.exists():
                                    respaldorecord = RespaldoRecordAcademico(
                                        recordacademicooriginal=recordantiguo,
                                        recordacademiconuevo=recordnuevo
                                    )
                                    respaldorecord.save()
                                else:
                                    respaldorecord = respaldo[0]
                                    respaldorecord.recordacademiconuevo = recordnuevo
                                    respaldorecord.save()
                                print(u"Record actualizado %s" % recordnuevo)

                        else:
                            hojadestino.write(fila, 3, "NO ENCONTRO RECORD ANTIGUO %s" % equivalencia.asignaturamalla)
                            fila += 1

                    practicasppf = inscripcion.numero_horas_practicas_pre_profesionales()
                    hojadestino.write(fila, 3, practicasppf, fuentenormal)
                    horasvinculacionf = inscripcion.numero_horas_proyectos_vinculacion()
                    hojadestino.write(fila, 4, horasvinculacionf, fuentenormal)
                    fila += 1

                    time.sleep(1)

            lin += 1

        libdestino.save(output_folder + libre_origen)
        print(output_folder + libre_origen)
        print("Proceso finalizado. . .")

    except Exception as ex:
        transaction.set_rollback(True)
        print('error: %s' % ex)
        hojadestino.write(fila, 3, str(ex))
        fila += 1
def homologacion_tics_7():
    # verificar que todos los estudiantes tengan la misma malla
    cadena = ''
    linea, excluidos, conexito = 0, 0, 0
    try:
        libre_origen = '/homologacion_tics_7.xls'
        fuentecabecera = easyxf(
            'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
        fuentenormal = easyxf(
            'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')

        output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
        # liborigen = xlrd.open_workbook(output_folder + libre_origen)
        libdestino = xlwt.Workbook()
        hojadestino = libdestino.add_sheet('Sheet1')
        fil = 0
        columnas = [(u"CEDULA", 6000),
                    (u"APELLIDOS Y NOMBRES", 6000),
                    (u"OBSERVACIÓN", 6000),
                    (u"HORAS PRACTICAS", 6000),
                    (u"HORAS VINCULACION", 6000),
                    (u"OBSERVACIÓN", 6000)
                    ]
        for col_num in range(len(columnas)):
            hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
            hojadestino.col(col_num).width = columnas[col_num][1]
        fila = 1
        titulacion = 0

        lin = 0
        # miarchivo = openpyxl.load_workbook("CEDULA_COMUNICACION_final.xlsx")
        miarchivo = openpyxl.load_workbook("tecnologialinea.xlsx")
        # miarchivo = openpyxl.load_workbook("prueba_2.xlsx")

        ws = miarchivo.get_sheet_by_name("septimo")
        worksheet = ws
        c = 0
        cont = 0
        periodo_id = 224
        carrera_id = 133
        mallaantigua_id = 202
        mallanueva_id = 478

        for row in worksheet.iter_rows(min_row=0):
            if lin >= 0:
                currentValues, cadena = [], ''
                for cell in row:
                    cadena += str(cell.value) + ' '
                    currentValues.append(str(cell.value))
                identificacion = currentValues[0]

                if not identificacion:
                    break

                with transaction.atomic():
                    matricula = Matricula.objects.filter(status=True, inscripcion__carrera__id=carrera_id,
                                                         inscripcion__persona__cedula=identificacion).first()
                    cont += 1
                    matricula.pasoayuda = True
                    matricula.save()
                    print(u"%s - %s" % (matricula, cont))
                    inscripcion = matricula.inscripcion
                    hojadestino.write(fila, 0, matricula.inscripcion.persona.identificacion(), fuentenormal)
                    hojadestino.write(fila, 1, matricula.inscripcion.persona.nombre_completo(), fuentenormal)
                    hojadestino.write(fila, 2, matricula.nivel.periodo.nombre, fuentenormal)

                    practicaspp = haber_cumplido_horas_creditos_practicas_preprofesionales(inscripcion.id)
                    horasvinculacion = haber_cumplido_horas_creditos_vinculacion(inscripcion.id)

                    if InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                       malla_id=mallaantigua_id).exists():
                        imantigua = \
                        InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallaantigua_id)[
                            0]
                        imantigua.status = False
                        imantigua.save()
                        print(u"Desactiva antigua inscripcion -----------------------------")

                    if not InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                           malla_id=mallanueva_id).exists():
                        imnueva = InscripcionMalla(inscripcion=inscripcion, malla_id=mallanueva_id)
                        imnueva.save()
                        print(u"Crea nueva inscripcion -----------------------------")

                    equivalencias = TablaEquivalenciaAsignaturas.objects.filter(status=True,
                                                                                asignaturamalla__malla_id=mallaantigua_id).order_by(
                        'asignaturamallasalto__nivelmalla__orden')
                    cont_asig_vinculacion_aprobadas = 0
                    horasfalta = 0
                    fechainicioitinerario = None
                    fechafinitinerario = None
                    for equivalencia in equivalencias:
                        print(u"nueva - %s" % equivalencia.asignaturamallasalto)
                        recordantiguo = inscripcion.recordacademico_set.filter(status=True,
                                                                               asignaturamalla_id=equivalencia.asignaturamalla_id).first()

                        if recordantiguo:
                            print(u"anterior - %s" % equivalencia.asignaturamalla)
                            print(u"Record antiguo: %s" % recordantiguo)
                            recordnuevo = None
                            recordantiguo.status = False
                            recordantiguo.save(update_asignaturamalla=False)

                            if equivalencia.asignaturamallasalto_id in [10587, 10618, 10612]:
                                observaciones = recordantiguo.observaciones + " Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15"
                                homologada = True


                            else:
                                observaciones = recordantiguo.observaciones + " Migración con base a RESOLUCIÓN CGA-SO-9-2023-NO15 Y RESOLUCIÓN OCS-SO-18-2023-NO14"
                                homologada = False
                            if not RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                  asignaturamalla=equivalencia.asignaturamallasalto).exists():

                                recordnuevo = RecordAcademico(inscripcion=inscripcion,
                                                              matriculas=recordantiguo.matriculas,
                                                              asignaturamalla=equivalencia.asignaturamallasalto,
                                                              asignatura=equivalencia.asignaturamallasalto.asignatura,
                                                              asignaturaold_id=recordantiguo.asignatura.id,
                                                              nota=recordantiguo.nota,
                                                              asistencia=recordantiguo.asistencia,
                                                              sinasistencia=recordantiguo.sinasistencia,
                                                              fecha=recordantiguo.fecha,
                                                              noaplica=recordantiguo.noaplica,
                                                              aprobada=recordantiguo.aprobada,
                                                              convalidacion=recordantiguo.convalidacion,
                                                              pendiente=recordantiguo.pendiente,
                                                              creditos=equivalencia.asignaturamallasalto.creditos,
                                                              horas=equivalencia.asignaturamallasalto.horas,
                                                              valida=recordantiguo.valida,
                                                              validapromedio=recordantiguo.validapromedio,
                                                              observaciones=observaciones,
                                                              homologada=homologada,
                                                              materiaregular=recordantiguo.materiaregular,
                                                              materiacurso=None,
                                                              completonota=recordantiguo.completonota,
                                                              completoasistencia=recordantiguo.completoasistencia,
                                                              fechainicio=recordantiguo.fechainicio,
                                                              fechafin=recordantiguo.fechafin,
                                                              suficiencia=recordantiguo.suficiencia,
                                                              asignaturamallahistorico_id=recordantiguo.asignaturamalla.id,
                                                              reverso=False)
                                recordnuevo.save()
                                print(u"Crea nuevo record %s" % recordnuevo)


                            elif RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                asignaturamalla=equivalencia.asignaturamallasalto):
                                recordnuevo = RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                             asignaturamalla=equivalencia.asignaturamallasalto)[
                                    0]
                                recordnuevo.matriculas = recordantiguo.matriculas
                                recordnuevo.asignaturamalla = equivalencia.asignaturamallasalto
                                recordnuevo.asignatura = equivalencia.asignaturamallasalto.asignatura
                                recordnuevo.asignaturaold = recordantiguo.asignaturamalla.asignatura
                                recordnuevo.nota = recordantiguo.nota
                                recordnuevo.asistencia = recordantiguo.asistencia
                                recordnuevo.sinasistencia = recordantiguo.sinasistencia
                                recordnuevo.fecha = recordantiguo.fecha
                                recordnuevo.noaplica = recordantiguo.noaplica
                                recordnuevo.aprobada = recordantiguo.aprobada
                                recordnuevo.convalidacion = recordantiguo.convalidacion
                                recordnuevo.pendiente = recordantiguo.pendiente
                                recordnuevo.creditos = equivalencia.asignaturamallasalto.creditos
                                recordnuevo.horas = equivalencia.asignaturamallasalto.horas
                                recordnuevo.valida = recordantiguo.valida
                                recordnuevo.validapromedio = recordantiguo.validapromedio
                                recordnuevo.observaciones = observaciones
                                recordnuevo.homologada = homologada
                                recordnuevo.materiaregular = recordantiguo.materiaregular
                                recordnuevo.materiacurso = None
                                recordnuevo.completonota = recordantiguo.completonota
                                recordnuevo.completoasistencia = recordantiguo.completoasistencia
                                recordnuevo.fechainicio = recordantiguo.fechainicio
                                recordnuevo.fechafin = recordantiguo.fechafin
                                recordnuevo.suficiencia = recordantiguo.suficiencia
                                recordnuevo.asignaturamallahistorico = recordantiguo.asignaturamalla
                                recordnuevo.reverso = False
                                recordnuevo.save()

                            if recordnuevo:
                                historicos = HistoricoRecordAcademico.objects.filter(status=True,
                                                                                     recordacademico=recordantiguo).update(
                                    recordacademico=recordnuevo,
                                    creditos=recordnuevo.creditos,
                                    horas=recordnuevo.horas,
                                    homologada=recordnuevo.homologada)
                                respaldo = RespaldoRecordAcademico.objects.filter(status=True,
                                                                                  recordacademicooriginal=recordantiguo)

                                if not respaldo.exists():
                                    respaldorecord = RespaldoRecordAcademico(
                                        recordacademicooriginal=recordantiguo,
                                        recordacademiconuevo=recordnuevo
                                    )
                                    respaldorecord.save()
                                else:
                                    respaldorecord = respaldo[0]
                                    respaldorecord.recordacademiconuevo = recordnuevo
                                    respaldorecord.save()
                                print(u"Record actualizado %s" % recordnuevo)

                        else:
                            hojadestino.write(fila, 3, "NO ENCONTRO RECORD ANTIGUO %s" % equivalencia.asignaturamalla)
                            fila += 1

                    practicasppf = inscripcion.numero_horas_practicas_pre_profesionales()
                    hojadestino.write(fila, 3, practicasppf, fuentenormal)
                    horasvinculacionf = inscripcion.numero_horas_proyectos_vinculacion()
                    hojadestino.write(fila, 4, horasvinculacionf, fuentenormal)
                    fila += 1

                    time.sleep(1)

            lin += 1

        libdestino.save(output_folder + libre_origen)
        print(output_folder + libre_origen)
        print("Proceso finalizado. . .")

    except Exception as ex:
        transaction.set_rollback(True)
        print('error: %s' % ex)
        hojadestino.write(fila, 3, str(ex))
        fila += 1
def homologacion_tics_8():
    # verificar que todos los estudiantes tengan la misma malla
    cadena = ''
    linea, excluidos, conexito = 0, 0, 0
    try:
        libre_origen = '/homologacion_tics_8.xls'
        fuentecabecera = easyxf(
            'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
        fuentenormal = easyxf(
            'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')

        output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
        # liborigen = xlrd.open_workbook(output_folder + libre_origen)
        libdestino = xlwt.Workbook()
        hojadestino = libdestino.add_sheet('Sheet1')
        fil = 0
        columnas = [(u"CEDULA", 6000),
                    (u"APELLIDOS Y NOMBRES", 6000),
                    (u"OBSERVACIÓN", 6000),
                    (u"HORAS PRACTICAS", 6000),
                    (u"HORAS VINCULACION", 6000),
                    (u"OBSERVACIÓN", 6000)
                    ]
        for col_num in range(len(columnas)):
            hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
            hojadestino.col(col_num).width = columnas[col_num][1]
        fila = 1
        titulacion = 0

        lin = 0
        # miarchivo = openpyxl.load_workbook("CEDULA_COMUNICACION_final.xlsx")
        miarchivo = openpyxl.load_workbook("tecnologialinea.xlsx")
        # miarchivo = openpyxl.load_workbook("prueba_2.xlsx")

        ws = miarchivo.get_sheet_by_name("octavo")
        worksheet = ws
        c = 0
        cont = 0
        periodo_id = 224
        carrera_id = 133
        mallaantigua_id = 202
        mallanueva_id = 478

        for row in worksheet.iter_rows(min_row=0):
            if lin >= 0:
                currentValues, cadena = [], ''
                for cell in row:
                    cadena += str(cell.value) + ' '
                    currentValues.append(str(cell.value))
                identificacion = currentValues[0]

                if not identificacion:
                    break

                with transaction.atomic():
                    matricula = Matricula.objects.filter(status=True, inscripcion__carrera__id=carrera_id,
                                                         inscripcion__persona__cedula=identificacion).first()
                    cont += 1
                    matricula.pasoayuda = True
                    matricula.save()
                    print(u"%s - %s" % (matricula, cont))
                    inscripcion = matricula.inscripcion
                    hojadestino.write(fila, 0, matricula.inscripcion.persona.identificacion(), fuentenormal)
                    hojadestino.write(fila, 1, matricula.inscripcion.persona.nombre_completo(), fuentenormal)
                    hojadestino.write(fila, 2, matricula.nivel.periodo.nombre, fuentenormal)

                    practicaspp = haber_cumplido_horas_creditos_practicas_preprofesionales(inscripcion.id)
                    horasvinculacion = haber_cumplido_horas_creditos_vinculacion(inscripcion.id)

                    if InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                       malla_id=mallaantigua_id).exists():
                        imantigua = \
                        InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallaantigua_id)[
                            0]
                        imantigua.status = False
                        imantigua.save()
                        print(u"Desactiva antigua inscripcion -----------------------------")

                    if not InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                           malla_id=mallanueva_id).exists():
                        imnueva = InscripcionMalla(inscripcion=inscripcion, malla_id=mallanueva_id)
                        imnueva.save()
                        print(u"Crea nueva inscripcion -----------------------------")

                    equivalencias = TablaEquivalenciaAsignaturas.objects.filter(status=True,
                                                                                asignaturamalla__malla_id=mallaantigua_id).order_by(
                        'asignaturamallasalto__nivelmalla__orden')
                    cont_asig_vinculacion_aprobadas = 0
                    horasfalta = 0
                    fechainicioitinerario = None
                    fechafinitinerario = None
                    for equivalencia in equivalencias:
                        print(u"nueva - %s" % equivalencia.asignaturamallasalto)
                        recordantiguo = inscripcion.recordacademico_set.filter(status=True,
                                                                               asignaturamalla_id=equivalencia.asignaturamalla_id).first()

                        if recordantiguo:
                            print(u"anterior - %s" % equivalencia.asignaturamalla)
                            print(u"Record antiguo: %s" % recordantiguo)
                            recordnuevo = None
                            recordantiguo.status = False
                            recordantiguo.save(update_asignaturamalla=False)

                            if equivalencia.asignaturamallasalto_id in [10587, 10618, 10612]:
                                observaciones = recordantiguo.observaciones + " Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15"
                                homologada = True


                            else:
                                observaciones = recordantiguo.observaciones + " Migración con base a RESOLUCIÓN CGA-SO-9-2023-NO15 Y RESOLUCIÓN OCS-SO-18-2023-NO14"
                                homologada = False
                            if not RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                  asignaturamalla=equivalencia.asignaturamallasalto).exists():

                                recordnuevo = RecordAcademico(inscripcion=inscripcion,
                                                              matriculas=recordantiguo.matriculas,
                                                              asignaturamalla=equivalencia.asignaturamallasalto,
                                                              asignatura=equivalencia.asignaturamallasalto.asignatura,
                                                              asignaturaold_id=recordantiguo.asignatura.id,
                                                              nota=recordantiguo.nota,
                                                              asistencia=recordantiguo.asistencia,
                                                              sinasistencia=recordantiguo.sinasistencia,
                                                              fecha=recordantiguo.fecha,
                                                              noaplica=recordantiguo.noaplica,
                                                              aprobada=recordantiguo.aprobada,
                                                              convalidacion=recordantiguo.convalidacion,
                                                              pendiente=recordantiguo.pendiente,
                                                              creditos=equivalencia.asignaturamallasalto.creditos,
                                                              horas=equivalencia.asignaturamallasalto.horas,
                                                              valida=recordantiguo.valida,
                                                              validapromedio=recordantiguo.validapromedio,
                                                              observaciones=observaciones,
                                                              homologada=homologada,
                                                              materiaregular=recordantiguo.materiaregular,
                                                              materiacurso=None,
                                                              completonota=recordantiguo.completonota,
                                                              completoasistencia=recordantiguo.completoasistencia,
                                                              fechainicio=recordantiguo.fechainicio,
                                                              fechafin=recordantiguo.fechafin,
                                                              suficiencia=recordantiguo.suficiencia,
                                                              asignaturamallahistorico_id=recordantiguo.asignaturamalla.id,
                                                              reverso=False)
                                recordnuevo.save()
                                print(u"Crea nuevo record %s" % recordnuevo)


                            elif RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                asignaturamalla=equivalencia.asignaturamallasalto):
                                recordnuevo = RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                             asignaturamalla=equivalencia.asignaturamallasalto)[
                                    0]
                                recordnuevo.matriculas = recordantiguo.matriculas
                                recordnuevo.asignaturamalla = equivalencia.asignaturamallasalto
                                recordnuevo.asignatura = equivalencia.asignaturamallasalto.asignatura
                                recordnuevo.asignaturaold = recordantiguo.asignaturamalla.asignatura
                                recordnuevo.nota = recordantiguo.nota
                                recordnuevo.asistencia = recordantiguo.asistencia
                                recordnuevo.sinasistencia = recordantiguo.sinasistencia
                                recordnuevo.fecha = recordantiguo.fecha
                                recordnuevo.noaplica = recordantiguo.noaplica
                                recordnuevo.aprobada = recordantiguo.aprobada
                                recordnuevo.convalidacion = recordantiguo.convalidacion
                                recordnuevo.pendiente = recordantiguo.pendiente
                                recordnuevo.creditos = equivalencia.asignaturamallasalto.creditos
                                recordnuevo.horas = equivalencia.asignaturamallasalto.horas
                                recordnuevo.valida = recordantiguo.valida
                                recordnuevo.validapromedio = recordantiguo.validapromedio
                                recordnuevo.observaciones = observaciones
                                recordnuevo.homologada = homologada
                                recordnuevo.materiaregular = recordantiguo.materiaregular
                                recordnuevo.materiacurso = None
                                recordnuevo.completonota = recordantiguo.completonota
                                recordnuevo.completoasistencia = recordantiguo.completoasistencia
                                recordnuevo.fechainicio = recordantiguo.fechainicio
                                recordnuevo.fechafin = recordantiguo.fechafin
                                recordnuevo.suficiencia = recordantiguo.suficiencia
                                recordnuevo.asignaturamallahistorico = recordantiguo.asignaturamalla
                                recordnuevo.reverso = False
                                recordnuevo.save()

                            if recordnuevo:
                                historicos = HistoricoRecordAcademico.objects.filter(status=True,
                                                                                     recordacademico=recordantiguo).update(
                                    recordacademico=recordnuevo,
                                    creditos=recordnuevo.creditos,
                                    horas=recordnuevo.horas,
                                    homologada=recordnuevo.homologada)
                                respaldo = RespaldoRecordAcademico.objects.filter(status=True,
                                                                                  recordacademicooriginal=recordantiguo)

                                if not respaldo.exists():
                                    respaldorecord = RespaldoRecordAcademico(
                                        recordacademicooriginal=recordantiguo,
                                        recordacademiconuevo=recordnuevo
                                    )
                                    respaldorecord.save()
                                else:
                                    respaldorecord = respaldo[0]
                                    respaldorecord.recordacademiconuevo = recordnuevo
                                    respaldorecord.save()
                                print(u"Record actualizado %s" % recordnuevo)

                        else:
                            hojadestino.write(fila, 3, "NO ENCONTRO RECORD ANTIGUO %s" % equivalencia.asignaturamalla)
                            fila += 1

                    practicasppf = inscripcion.numero_horas_practicas_pre_profesionales()
                    hojadestino.write(fila, 3, practicasppf, fuentenormal)
                    horasvinculacionf = inscripcion.numero_horas_proyectos_vinculacion()
                    hojadestino.write(fila, 4, horasvinculacionf, fuentenormal)
                    fila += 1

                    time.sleep(1)

            lin += 1

        libdestino.save(output_folder + libre_origen)
        print(output_folder + libre_origen)
        print("Proceso finalizado. . .")

    except Exception as ex:
        transaction.set_rollback(True)
        print('error: %s' % ex)
        hojadestino.write(fila, 3, str(ex))
        fila += 1
def homologacion_tics_9():
    # verificar que todos los estudiantes tengan la misma malla
    cadena = ''
    linea, excluidos, conexito = 0, 0, 0
    try:
        libre_origen = '/homologacion_tics_9.xls'
        fuentecabecera = easyxf(
            'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
        fuentenormal = easyxf(
            'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')

        output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
        # liborigen = xlrd.open_workbook(output_folder + libre_origen)
        libdestino = xlwt.Workbook()
        hojadestino = libdestino.add_sheet('Sheet1')
        fil = 0
        columnas = [(u"CEDULA", 6000),
                    (u"APELLIDOS Y NOMBRES", 6000),
                    (u"OBSERVACIÓN", 6000),
                    (u"HORAS PRACTICAS", 6000),
                    (u"HORAS VINCULACION", 6000),
                    (u"OBSERVACIÓN", 6000)
                    ]
        for col_num in range(len(columnas)):
            hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
            hojadestino.col(col_num).width = columnas[col_num][1]
        fila = 1
        titulacion = 0

        lin = 0
        # miarchivo = openpyxl.load_workbook("CEDULA_COMUNICACION_final.xlsx")
        miarchivo = openpyxl.load_workbook("tecnologialinea.xlsx")
        # miarchivo = openpyxl.load_workbook("prueba_2.xlsx")

        ws = miarchivo.get_sheet_by_name("noveno")
        worksheet = ws
        c = 0
        cont = 0
        periodo_id = 224
        carrera_id = 133
        mallaantigua_id = 202
        mallanueva_id = 478

        for row in worksheet.iter_rows(min_row=0):
            if lin >= 0:
                currentValues, cadena = [], ''
                for cell in row:
                    cadena += str(cell.value) + ' '
                    currentValues.append(str(cell.value))
                identificacion = currentValues[0]

                if not identificacion:
                    break

                with transaction.atomic():
                    matricula = Matricula.objects.filter(status=True, inscripcion__carrera__id=carrera_id,
                                                         inscripcion__persona__cedula=identificacion).first()
                    cont += 1
                    matricula.pasoayuda = True
                    matricula.save()
                    print(u"%s - %s" % (matricula, cont))
                    inscripcion = matricula.inscripcion
                    hojadestino.write(fila, 0, matricula.inscripcion.persona.identificacion(), fuentenormal)
                    hojadestino.write(fila, 1, matricula.inscripcion.persona.nombre_completo(), fuentenormal)
                    hojadestino.write(fila, 2, matricula.nivel.periodo.nombre, fuentenormal)

                    practicaspp = haber_cumplido_horas_creditos_practicas_preprofesionales(inscripcion.id)
                    horasvinculacion = haber_cumplido_horas_creditos_vinculacion(inscripcion.id)

                    if InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                       malla_id=mallaantigua_id).exists():
                        imantigua = \
                        InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallaantigua_id)[
                            0]
                        imantigua.status = False
                        imantigua.save()
                        print(u"Desactiva antigua inscripcion -----------------------------")

                    if not InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                           malla_id=mallanueva_id).exists():
                        imnueva = InscripcionMalla(inscripcion=inscripcion, malla_id=mallanueva_id)
                        imnueva.save()
                        print(u"Crea nueva inscripcion -----------------------------")

                    equivalencias = TablaEquivalenciaAsignaturas.objects.filter(status=True,
                                                                                asignaturamalla__malla_id=mallaantigua_id).order_by(
                        'asignaturamallasalto__nivelmalla__orden')
                    cont_asig_vinculacion_aprobadas = 0
                    horasfalta = 0
                    fechainicioitinerario = None
                    fechafinitinerario = None
                    for equivalencia in equivalencias:
                        print(u"nueva - %s" % equivalencia.asignaturamallasalto)
                        recordantiguo = inscripcion.recordacademico_set.filter(status=True,
                                                                               asignaturamalla_id=equivalencia.asignaturamalla_id).first()

                        if recordantiguo:
                            print(u"anterior - %s" % equivalencia.asignaturamalla)
                            print(u"Record antiguo: %s" % recordantiguo)
                            recordnuevo = None
                            recordantiguo.status = False
                            recordantiguo.save(update_asignaturamalla=False)

                            if equivalencia.asignaturamallasalto_id in [10587, 10618, 10612]:
                                observaciones = recordantiguo.observaciones + " Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15"
                                homologada = True


                            else:
                                observaciones = recordantiguo.observaciones + " Migración con base a RESOLUCIÓN CGA-SO-9-2023-NO15 Y RESOLUCIÓN OCS-SO-18-2023-NO14"
                                homologada = False
                            if not RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                  asignaturamalla=equivalencia.asignaturamallasalto).exists():

                                recordnuevo = RecordAcademico(inscripcion=inscripcion,
                                                              matriculas=recordantiguo.matriculas,
                                                              asignaturamalla=equivalencia.asignaturamallasalto,
                                                              asignatura=equivalencia.asignaturamallasalto.asignatura,
                                                              asignaturaold_id=recordantiguo.asignatura.id,
                                                              nota=recordantiguo.nota,
                                                              asistencia=recordantiguo.asistencia,
                                                              sinasistencia=recordantiguo.sinasistencia,
                                                              fecha=recordantiguo.fecha,
                                                              noaplica=recordantiguo.noaplica,
                                                              aprobada=recordantiguo.aprobada,
                                                              convalidacion=recordantiguo.convalidacion,
                                                              pendiente=recordantiguo.pendiente,
                                                              creditos=equivalencia.asignaturamallasalto.creditos,
                                                              horas=equivalencia.asignaturamallasalto.horas,
                                                              valida=recordantiguo.valida,
                                                              validapromedio=recordantiguo.validapromedio,
                                                              observaciones=observaciones,
                                                              homologada=homologada,
                                                              materiaregular=recordantiguo.materiaregular,
                                                              materiacurso=None,
                                                              completonota=recordantiguo.completonota,
                                                              completoasistencia=recordantiguo.completoasistencia,
                                                              fechainicio=recordantiguo.fechainicio,
                                                              fechafin=recordantiguo.fechafin,
                                                              suficiencia=recordantiguo.suficiencia,
                                                              asignaturamallahistorico_id=recordantiguo.asignaturamalla.id,
                                                              reverso=False)
                                recordnuevo.save()
                                print(u"Crea nuevo record %s" % recordnuevo)


                            elif RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                asignaturamalla=equivalencia.asignaturamallasalto):
                                recordnuevo = RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                             asignaturamalla=equivalencia.asignaturamallasalto)[
                                    0]
                                recordnuevo.matriculas = recordantiguo.matriculas
                                recordnuevo.asignaturamalla = equivalencia.asignaturamallasalto
                                recordnuevo.asignatura = equivalencia.asignaturamallasalto.asignatura
                                recordnuevo.asignaturaold = recordantiguo.asignaturamalla.asignatura
                                recordnuevo.nota = recordantiguo.nota
                                recordnuevo.asistencia = recordantiguo.asistencia
                                recordnuevo.sinasistencia = recordantiguo.sinasistencia
                                recordnuevo.fecha = recordantiguo.fecha
                                recordnuevo.noaplica = recordantiguo.noaplica
                                recordnuevo.aprobada = recordantiguo.aprobada
                                recordnuevo.convalidacion = recordantiguo.convalidacion
                                recordnuevo.pendiente = recordantiguo.pendiente
                                recordnuevo.creditos = equivalencia.asignaturamallasalto.creditos
                                recordnuevo.horas = equivalencia.asignaturamallasalto.horas
                                recordnuevo.valida = recordantiguo.valida
                                recordnuevo.validapromedio = recordantiguo.validapromedio
                                recordnuevo.observaciones = observaciones
                                recordnuevo.homologada = homologada
                                recordnuevo.materiaregular = recordantiguo.materiaregular
                                recordnuevo.materiacurso = None
                                recordnuevo.completonota = recordantiguo.completonota
                                recordnuevo.completoasistencia = recordantiguo.completoasistencia
                                recordnuevo.fechainicio = recordantiguo.fechainicio
                                recordnuevo.fechafin = recordantiguo.fechafin
                                recordnuevo.suficiencia = recordantiguo.suficiencia
                                recordnuevo.asignaturamallahistorico = recordantiguo.asignaturamalla
                                recordnuevo.reverso = False
                                recordnuevo.save()

                            if recordnuevo:
                                historicos = HistoricoRecordAcademico.objects.filter(status=True,
                                                                                     recordacademico=recordantiguo).update(
                                    recordacademico=recordnuevo,
                                    creditos=recordnuevo.creditos,
                                    horas=recordnuevo.horas,
                                    homologada=recordnuevo.homologada)
                                respaldo = RespaldoRecordAcademico.objects.filter(status=True,
                                                                                  recordacademicooriginal=recordantiguo)

                                if not respaldo.exists():
                                    respaldorecord = RespaldoRecordAcademico(
                                        recordacademicooriginal=recordantiguo,
                                        recordacademiconuevo=recordnuevo
                                    )
                                    respaldorecord.save()
                                else:
                                    respaldorecord = respaldo[0]
                                    respaldorecord.recordacademiconuevo = recordnuevo
                                    respaldorecord.save()
                                print(u"Record actualizado %s" % recordnuevo)

                        else:
                            hojadestino.write(fila, 3, "NO ENCONTRO RECORD ANTIGUO %s" % equivalencia.asignaturamalla)
                            fila += 1

                    practicasppf = inscripcion.numero_horas_practicas_pre_profesionales()
                    hojadestino.write(fila, 3, practicasppf, fuentenormal)
                    horasvinculacionf = inscripcion.numero_horas_proyectos_vinculacion()
                    hojadestino.write(fila, 4, horasvinculacionf, fuentenormal)
                    fila += 1

                    time.sleep(1)

            lin += 1

        libdestino.save(output_folder + libre_origen)
        print(output_folder + libre_origen)
        print("Proceso finalizado. . .")

    except Exception as ex:
        transaction.set_rollback(True)
        print('error: %s' % ex)
        hojadestino.write(fila, 3, str(ex))
        fila += 1
def homologacion_tics_10():
    # verificar que todos los estudiantes tengan la misma malla
    cadena = ''
    linea, excluidos, conexito = 0, 0, 0
    try:
        libre_origen = '/homologacion_tics_10.xls'
        fuentecabecera = easyxf(
            'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
        fuentenormal = easyxf(
            'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')

        output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
        # liborigen = xlrd.open_workbook(output_folder + libre_origen)
        libdestino = xlwt.Workbook()
        hojadestino = libdestino.add_sheet('Sheet1')
        fil = 0
        columnas = [(u"CEDULA", 6000),
                    (u"APELLIDOS Y NOMBRES", 6000),
                    (u"OBSERVACIÓN", 6000),
                    (u"HORAS PRACTICAS", 6000),
                    (u"HORAS VINCULACION", 6000),
                    (u"OBSERVACIÓN", 6000)
                    ]
        for col_num in range(len(columnas)):
            hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
            hojadestino.col(col_num).width = columnas[col_num][1]
        fila = 1
        titulacion = 0

        lin = 0
        # miarchivo = openpyxl.load_workbook("CEDULA_COMUNICACION_final.xlsx")
        miarchivo = openpyxl.load_workbook("tecnologialinea.xlsx")
        # miarchivo = openpyxl.load_workbook("prueba_2.xlsx")

        ws = miarchivo.get_sheet_by_name("decimo")
        worksheet = ws
        c = 0
        cont = 0
        periodo_id = 224
        carrera_id = 133
        mallaantigua_id = 202
        mallanueva_id = 478

        for row in worksheet.iter_rows(min_row=0):
            if lin >= 0:
                currentValues, cadena = [], ''
                for cell in row:
                    cadena += str(cell.value) + ' '
                    currentValues.append(str(cell.value))
                identificacion = currentValues[0]

                if not identificacion:
                    break

                with transaction.atomic():
                    matricula = Matricula.objects.filter(status=True, inscripcion__carrera__id=carrera_id,
                                                         inscripcion__persona__cedula=identificacion).first()
                    cont += 1
                    matricula.pasoayuda = True
                    matricula.save()
                    print(u"%s - %s" % (matricula, cont))
                    inscripcion = matricula.inscripcion
                    hojadestino.write(fila, 0, matricula.inscripcion.persona.identificacion(), fuentenormal)
                    hojadestino.write(fila, 1, matricula.inscripcion.persona.nombre_completo(), fuentenormal)
                    hojadestino.write(fila, 2, matricula.nivel.periodo.nombre, fuentenormal)

                    practicaspp = haber_cumplido_horas_creditos_practicas_preprofesionales(inscripcion.id)
                    horasvinculacion = haber_cumplido_horas_creditos_vinculacion(inscripcion.id)

                    if InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                       malla_id=mallaantigua_id).exists():
                        imantigua = \
                        InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallaantigua_id)[
                            0]
                        imantigua.status = False
                        imantigua.save()
                        print(u"Desactiva antigua inscripcion -----------------------------")

                    if not InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                           malla_id=mallanueva_id).exists():
                        imnueva = InscripcionMalla(inscripcion=inscripcion, malla_id=mallanueva_id)
                        imnueva.save()
                        print(u"Crea nueva inscripcion -----------------------------")

                    equivalencias = TablaEquivalenciaAsignaturas.objects.filter(status=True,
                                                                                asignaturamalla__malla_id=mallaantigua_id).order_by(
                        'asignaturamallasalto__nivelmalla__orden')
                    cont_asig_vinculacion_aprobadas = 0
                    horasfalta = 0
                    fechainicioitinerario = None
                    fechafinitinerario = None
                    for equivalencia in equivalencias:
                        print(u"nueva - %s" % equivalencia.asignaturamallasalto)
                        recordantiguo = inscripcion.recordacademico_set.filter(status=True,
                                                                               asignaturamalla_id=equivalencia.asignaturamalla_id).first()

                        if recordantiguo:
                            print(u"anterior - %s" % equivalencia.asignaturamalla)
                            print(u"Record antiguo: %s" % recordantiguo)
                            recordnuevo = None
                            recordantiguo.status = False
                            recordantiguo.save(update_asignaturamalla=False)

                            if equivalencia.asignaturamallasalto_id in [10587, 10618, 10612]:
                                observaciones = recordantiguo.observaciones + " Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15"
                                homologada = True


                            else:
                                observaciones = recordantiguo.observaciones + " Migración con base a RESOLUCIÓN CGA-SO-9-2023-NO15 Y RESOLUCIÓN OCS-SO-18-2023-NO14"
                                homologada = False
                            if not RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                  asignaturamalla=equivalencia.asignaturamallasalto).exists():

                                recordnuevo = RecordAcademico(inscripcion=inscripcion,
                                                              matriculas=recordantiguo.matriculas,
                                                              asignaturamalla=equivalencia.asignaturamallasalto,
                                                              asignatura=equivalencia.asignaturamallasalto.asignatura,
                                                              asignaturaold_id=recordantiguo.asignatura.id,
                                                              nota=recordantiguo.nota,
                                                              asistencia=recordantiguo.asistencia,
                                                              sinasistencia=recordantiguo.sinasistencia,
                                                              fecha=recordantiguo.fecha,
                                                              noaplica=recordantiguo.noaplica,
                                                              aprobada=recordantiguo.aprobada,
                                                              convalidacion=recordantiguo.convalidacion,
                                                              pendiente=recordantiguo.pendiente,
                                                              creditos=equivalencia.asignaturamallasalto.creditos,
                                                              horas=equivalencia.asignaturamallasalto.horas,
                                                              valida=recordantiguo.valida,
                                                              validapromedio=recordantiguo.validapromedio,
                                                              observaciones=observaciones,
                                                              homologada=homologada,
                                                              materiaregular=recordantiguo.materiaregular,
                                                              materiacurso=None,
                                                              completonota=recordantiguo.completonota,
                                                              completoasistencia=recordantiguo.completoasistencia,
                                                              fechainicio=recordantiguo.fechainicio,
                                                              fechafin=recordantiguo.fechafin,
                                                              suficiencia=recordantiguo.suficiencia,
                                                              asignaturamallahistorico_id=recordantiguo.asignaturamalla.id,
                                                              reverso=False)
                                recordnuevo.save()
                                print(u"Crea nuevo record %s" % recordnuevo)


                            elif RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                asignaturamalla=equivalencia.asignaturamallasalto):
                                recordnuevo = RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                             asignaturamalla=equivalencia.asignaturamallasalto)[
                                    0]
                                recordnuevo.matriculas = recordantiguo.matriculas
                                recordnuevo.asignaturamalla = equivalencia.asignaturamallasalto
                                recordnuevo.asignatura = equivalencia.asignaturamallasalto.asignatura
                                recordnuevo.asignaturaold = recordantiguo.asignaturamalla.asignatura
                                recordnuevo.nota = recordantiguo.nota
                                recordnuevo.asistencia = recordantiguo.asistencia
                                recordnuevo.sinasistencia = recordantiguo.sinasistencia
                                recordnuevo.fecha = recordantiguo.fecha
                                recordnuevo.noaplica = recordantiguo.noaplica
                                recordnuevo.aprobada = recordantiguo.aprobada
                                recordnuevo.convalidacion = recordantiguo.convalidacion
                                recordnuevo.pendiente = recordantiguo.pendiente
                                recordnuevo.creditos = equivalencia.asignaturamallasalto.creditos
                                recordnuevo.horas = equivalencia.asignaturamallasalto.horas
                                recordnuevo.valida = recordantiguo.valida
                                recordnuevo.validapromedio = recordantiguo.validapromedio
                                recordnuevo.observaciones = observaciones
                                recordnuevo.homologada = homologada
                                recordnuevo.materiaregular = recordantiguo.materiaregular
                                recordnuevo.materiacurso = None
                                recordnuevo.completonota = recordantiguo.completonota
                                recordnuevo.completoasistencia = recordantiguo.completoasistencia
                                recordnuevo.fechainicio = recordantiguo.fechainicio
                                recordnuevo.fechafin = recordantiguo.fechafin
                                recordnuevo.suficiencia = recordantiguo.suficiencia
                                recordnuevo.asignaturamallahistorico = recordantiguo.asignaturamalla
                                recordnuevo.reverso = False
                                recordnuevo.save()

                            if recordnuevo:
                                historicos = HistoricoRecordAcademico.objects.filter(status=True,
                                                                                     recordacademico=recordantiguo).update(
                                    recordacademico=recordnuevo,
                                    creditos=recordnuevo.creditos,
                                    horas=recordnuevo.horas,
                                    homologada=recordnuevo.homologada)
                                respaldo = RespaldoRecordAcademico.objects.filter(status=True,
                                                                                  recordacademicooriginal=recordantiguo)

                                if not respaldo.exists():
                                    respaldorecord = RespaldoRecordAcademico(
                                        recordacademicooriginal=recordantiguo,
                                        recordacademiconuevo=recordnuevo
                                    )
                                    respaldorecord.save()
                                else:
                                    respaldorecord = respaldo[0]
                                    respaldorecord.recordacademiconuevo = recordnuevo
                                    respaldorecord.save()
                                print(u"Record actualizado %s" % recordnuevo)

                        else:
                            hojadestino.write(fila, 3, "NO ENCONTRO RECORD ANTIGUO %s" % equivalencia.asignaturamalla)
                            fila += 1

                    practicasppf = inscripcion.numero_horas_practicas_pre_profesionales()
                    hojadestino.write(fila, 3, practicasppf, fuentenormal)
                    horasvinculacionf = inscripcion.numero_horas_proyectos_vinculacion()
                    hojadestino.write(fila, 4, horasvinculacionf, fuentenormal)
                    fila += 1

                    time.sleep(1)

            lin += 1

        libdestino.save(output_folder + libre_origen)
        print(output_folder + libre_origen)
        print("Proceso finalizado. . .")

    except Exception as ex:
        transaction.set_rollback(True)
        print('error: %s' % ex)
        hojadestino.write(fila, 3, str(ex))
        fila += 1

print("Función varios")
with concurrent.futures.ThreadPoolExecutor(max_workers=10) as executor:
    from settings import DEBUG
    # if DEBUG:
    #     path_anexo = 'reporte_acortezl__examen_admisión_enero2024.xlsx'
    future_1 = executor.submit(homologacion_tics)
    future_2 = executor.submit(homologacion_tics_2)
    future_3 = executor.submit(homologacion_tics_3)
    future_4 = executor.submit(homologacion_tics_4)
    future_5 = executor.submit(homologacion_tics_5)
    future_6 = executor.submit(homologacion_tics_6)
    future_7 = executor.submit(homologacion_tics_7)
    future_8 = executor.submit(homologacion_tics_8)
    future_9 = executor.submit(homologacion_tics_9)
    future_10 = executor.submit(homologacion_tics_10)