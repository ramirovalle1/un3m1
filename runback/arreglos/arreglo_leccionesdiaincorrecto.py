#!/usr/bin/env python
# -*- coding: utf-8 -*-
import io
import os
import sys

# SITE_ROOT = os.path.dirname(os.path.realpath(__file__))
from openpyxl import load_workbook
YOUR_PATH = os.path.dirname(os.path.realpath(__file__))
#print(f"YOUR_PATH: {YOUR_PATH}")
SITE_ROOT = os.path.dirname(os.path.dirname(YOUR_PATH))
SITE_ROOT = os.path.join(SITE_ROOT, '')
# print(f"SITE_ROOT: {SITE_ROOT}")
your_djangoproject_home = os.path.split(SITE_ROOT)[0]
# print(f"your_djangoproject_home: {your_djangoproject_home}")
sys.path.append(your_djangoproject_home)
from xlwt import *
from django.core.wsgi import get_wsgi_application

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "settings")
application = get_wsgi_application()
from django.db import transaction
from django.db.models import F, Q
from certi.models import Carnet
from sga.models import Periodo
from settings import SITE_STORAGE


def lecciones_dia_incorrecto():
    from sga.models import Leccion, DIAS_CHOICES
    font_style = XFStyle()
    font_style.font.bold = True
    font_style2 = XFStyle()
    font_style2.font.bold = False
    date_format = XFStyle()
    date_format.num_format_str = 'yyyy/mm/dd'
    archivo_ = 'LISTADO_LECCIONES_INCORRECTAS'
    url_archivo = "{}/media/{}.xls".format(SITE_STORAGE, archivo_)
    wb = Workbook(encoding='utf-8')
    ws = wb.add_sheet('LECCIONES_INCORRECTAS')
    # wb = load_workbook(filename=url_archivo, read_only=True)
    # ws = wb[wb.sheetnames[3]]
    columns = [
        (u"LECCION_ID", 2000),
        (u"CLASE_ID", 2000),
        (u"MATERIA_ID", 2000),
        (u"MATERIA", 25000),
        (u"DOCENTE", 17000),
        (u"TIPO DOCENTE", 8000),
        (u"FECHA", 6000),
        (u"# TURNO", 2000),
        (u"TURNO", 9000),
        (u"CLASE_DIA", 3000),
        (u"LECCION_DIA", 3000),
        (u"CLASE_NOMBRE_DIA", 8000),
        (u"LECCION_NOMBRE_DIA", 8000),
        (u"TIPO_HORARIO", 10000),
    ]
    row_num = 0
    for col_num in range(len(columns)):
        ws.write(row_num, col_num, columns[col_num][0], font_style)
        ws.col(col_num).width = columns[col_num][1]
    date_format = XFStyle()
    date_format.num_format_str = 'yyyy/mm/dd'
    row_num = 1
    lecciones = Leccion.objects.filter(~Q(fecha__iso_week_day=F('clase__dia')))

    for leccion in lecciones:
        ws.write(row_num, 0, leccion.id, font_style2)
        ws.write(row_num, 1, leccion.clase.id, font_style2)
        ws.write(row_num, 2, leccion.clase.materia.id, font_style2)
        ws.write(row_num, 3, leccion.clase.materia.__str__(), font_style2)
        ws.write(row_num, 4, leccion.clase.profesor.__str__(), font_style2)
        ws.write(row_num, 5, leccion.clase.tipoprofesor.__str__(), font_style2)
        ws.write(row_num, 6, leccion.fecha, date_format)
        ws.write(row_num, 7, leccion.clase.turno.turno, font_style2)
        ws.write(row_num, 8, f'{leccion.clase.turno.comienza} - {leccion.clase.turno.termina}', font_style2)
        ws.write(row_num, 9, leccion.clase.dia, font_style2)
        ws.write(row_num, 10, leccion.fecha.isoweekday(), font_style2)
        ws.write(row_num, 11, leccion.clase.get_dia_display(), font_style2)
        ws.write(row_num, 12, dict(DIAS_CHOICES)[leccion.fecha.isoweekday()], font_style2)
        ws.write(row_num, 13, leccion.clase.get_tipohorario_display(), font_style2)
        row_num += 1
    wb.save(url_archivo)
    print('Ruta del Archivo: ', url_archivo)

def correccion_lecciones_dia_incorrecto():
    from sga.models import Leccion, DIAS_CHOICES

    lecciones = Leccion.objects.filter(~Q(fecha__iso_week_day=F('clase__dia')), clase__materia__nivel__periodo_id=126)
    print('Numero de Registro', lecciones.count())
    for num, leccion in enumerate(lecciones):
        with transaction.atomic(using='default'):
            try:
                fecha = leccion.fecha
                fecha_correcta = leccion.fecha_clase_verbose()
                leccion_registrada = Leccion.objects.filter(clase=leccion.clase, fecha=fecha_correcta)
                print(leccion, leccion_registrada)


                # if not leccion_registrada.exists():
                #     leccion.fecha = fecha_correcta
                #     leccion.save()
                #     print(f'{num+1}.- {leccion.id} - {leccion} -- FECHA ACTUAL [{fecha}]  *****  FECHA CAMBIO ----> {fecha_correcta} ')
                #     for al in leccion.asistencialeccion_set.all():
                #         al.asistio=True#.update(asistio=True, virtual=True)
                #         al.virtual=True
                #         al.save()
                # else:
                #     leccion_duplicada = leccion_registrada.first()
                #     print(f'{num+1}.- {leccion.id} --- {leccion_duplicada.id} -- Leccion Incorrecta -- {leccion} ({leccion.fecha.strftime("%d-%m-%Y"), fecha_correcta.strftime("%d-%m-%Y")}) Leccion Correcta -----> {leccion_duplicada} ({leccion_duplicada.fecha}, {leccion_duplicada.fecha_clase_verbose()})')
                #     for al in leccion_duplicada.asistencialeccion_set.all():
                #         al.asistio=True#.update(asistio=True, virtual=True)
                #         al.virtual=True
                #         al.save()
            except Exception as ex:
                print(ex)


correccion_lecciones_dia_incorrecto()
# correccion_lecciones_dia_incorrecto()
