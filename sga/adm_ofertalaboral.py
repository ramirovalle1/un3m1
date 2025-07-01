# -*- coding: latin-1 -*-
import random
from datetime import datetime, timedelta

from openpyxl import workbook as openxl
from openpyxl.styles import Font as openxlFont
from openpyxl.styles.alignment import Alignment as alin

from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import Group
from django.db import transaction
from django.db.models import Q
from django.http import HttpResponseRedirect, JsonResponse, HttpResponse
from django.shortcuts import render
from django.template.loader import get_template

from decorators import last_access, secure_module
from empresa.commonviews import calculate_username_empresa
from empresa.models import RepresentantesEmpresa
from settings import EMAIL_DOMAIN
from sga.commonviews import adduserdata, obtener_reporte
from sga.forms import OfertaLaboralForm, AgendarCitaForm, AreaForm, \
    ConfiguracionForm, ObservacionesOfertaLaboralForm, GestionarOfertaForm, GestionarEmpresaForm, GrupoCarreraForm
from sga.funciones import MiPaginador, log, generar_usuario_cedula, notificacion, variable_valor
from sga.models import OfertaLaboral, AplicanteOferta, miinstitucion, \
    NIVEL_CONOCIMIENTO, CategoriaHerramienta, \
    ConocimientoInformatico, AreaOfertaLaboral, ConfiguraOferta, CUENTAS_CORREOS, AplicanteOfertaObservacion, Empleador, \
    User, PerfilUsuario, Graduado, Inscripcion, Persona, Notificacion, CarreraGrupo, Inscripcion
from sga.tasks import send_html_mail, conectar_cuenta
from empleo.models import OfertaLaboralEmpresa, SolicitudAprobacionEmpresa, HistorialAprobacionEmpresa, \
    PersonaAplicaOferta, ESTADOS_APLICACION, ESTADOS_EMPRESA, PersonaAplicaOferta, SolicitudRevisionTitulo
from sga.templatetags.sga_extras import encrypt
from bd.models import LogEntryLogin


@login_required(redirect_field_name='ret', login_url='/loginsga')
@secure_module
@last_access
@transaction.atomic()

#------------------------------------------Ver Listado-----------------------------------------------------------------
def view(request):
    global ex
    hoy = datetime.now().date()
    persona = request.session['persona']
    if request.method == 'POST':
        if 'action' in request.POST:
            action = request.POST['action']


            if action == 'verdescripcion':
                try:
                    oferta = OfertaLaboral.objects.get(pk=request.POST['id'])
                    return JsonResponse({'result': 'ok', 'area': oferta.area.descripcion, 'cargo': oferta.cargo, 'descripcion': oferta.descripcion})
                except Exception as ex:
                    return JsonResponse({'result': 'bad', "mensaje": u'Error al obtener los datos'})
#-----------------------------------------------------------------------------------------------------------
            #delcarreragrupo

            if action == 'deletegrupo':
                try:
                    carreragrupo = CarreraGrupo.objects.get(pk=encrypt(request.POST['id']))
                    carreragrupo.status = False
                    log(u'Elimino grupo de carrera: %s' % carreragrupo, request, "del")
                    carreragrupo.save(request)
                    return JsonResponse({"error": False})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"error": True, "mensaje": u"Error al eliminar los datos."})
#----------------------------------------------------------------------------------------------------------------------------


            if action == 'editgrupocar':
                try:
                    grucarrera = CarreraGrupo.objects.get(pk=encrypt(request.POST['id']))
                    f = GrupoCarreraForm(request.POST)
                    if f.is_valid():

                        grucarrera.nombre = f.cleaned_data['nombre']

                        estado = f.cleaned_data['status']
                        # 1 activo y 2 incactivo
                        b = True
                        if estado == '2':
                            b = False

                        grucarrera.activa = b

                        grucarrera.save(request)
                        log(u'Modifico area oferta laboral: %s' % grucarrera, request, "edit")
                        return JsonResponse({"result": "ok"})
                    else:
                        raise NameError('Error')
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})


            if action == 'agregargrupocarrera':
                try:

                    f = GrupoCarreraForm(request.POST)
                    if f.is_valid():

                        #lo que esta en f.cleaned_data es lo que recibe por post esta en el .form
                        if CarreraGrupo.objects.filter(nombre=f.cleaned_data['nombre'], status=True):
                            return JsonResponse({"result":"bad", "mensaje":u"Grupo carrera ya existe"})

                        estado= f.cleaned_data['status']

                        # 1 activo y 2 incactivo
                        b=True
                        if estado=='2':
                            b=False

                        grupocarrera=CarreraGrupo(nombre=f.cleaned_data['nombre'], activa=b)
                        grupocarrera.save(request)
                        log(u'adiciono grupo carrera: %s' % grupocarrera, request, "add")
                        return JsonResponse({'result': 'ok'})
                    else:
                        raise NameError('Error')

                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"error": True, "mensaje": u"Error al guardar los datos."})

            if action == 'asigcita':
                try:
                    registrado = AplicanteOferta.objects.get(pk=request.POST['id'])
                    f = AgendarCitaForm(request.POST)
                    if f.is_valid():
                        registrado.fechaentrevista = f.cleaned_data['fechaentrevista']
                        registrado.horaentrevista = f.cleaned_data['horaentrevista']
                        registrado.lugar = f.cleaned_data['lugar']
                        registrado.personacontacto = f.cleaned_data['personacontacto']
                        registrado.telefonocontacto = f.cleaned_data['telefonocontacto']
                        registrado.citaconfirmada = False
                        registrado.save(request)
                        log(u'Agendo cita de oferta laboral: %s' % registrado, request, "edit")
                        send_html_mail("Cita para oferta laboral aplicada", "emails/nuevacita.html", {'sistema': request.session['nombresistema'], 'registro': registrado, 't': miinstitucion(), 'dominio': EMAIL_DOMAIN}, registrado.inscripcion.persona.lista_emails_envio(), [])
                        return JsonResponse({"result": "ok"})
                    else:
                         raise NameError('Error')
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

            if action == 'del':
                try:
                    oferta = OfertaLaboral.objects.get(pk=request.POST['id'])
                    if oferta.aplicanteoferta_set.exists():
                        return JsonResponse({"result": "bad", "mensaje": u"Existen aplicaciones a esta oferta."})
                    log(u'Elimino oferta laboral: %s' % oferta, request, "del")
                    oferta.delete()
                    return JsonResponse({"result": "ok", "id": oferta.id})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

            if action == 'editsbu':
                try:
                    from sga.models import VariablesGlobales
                    if VariablesGlobales.objects.db_manager("sga_select").filter(status=True, variable__exact='SBU_VALOR').exists():
                        var = VariablesGlobales.objects.db_manager("sga_select").get(status=True, variable__exact='SBU_VALOR')
                        var.valor = request.POST['sbu_valor']
                        var.save(request)
                        log(u'Modifico SBU: %s' % var, request, "edit")
                    return JsonResponse({"result": False, 'mensaje': 'Edicion Exitosa'})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": True, "mensaje": u"Error al guardar los datos."})

            if action == 'delarea':
                try:
                    area = AreaOfertaLaboral.objects.get(pk=request.POST['id'])
                    area.status = False
                    log(u'Elimino area oferta laboral: %s' % area, request, "del")
                    area.save(request)
                    return JsonResponse({"result": "ok"})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

            if action == 'delpersonaresponsable':
                try:
                    personaresponable = ConfiguraOferta.objects.get(pk=request.POST['id'])
                    personaresponable.status = False
                    log(u'Elimino persona responsable: %s' % personaresponable, request, "del")
                    personaresponable.save(request)
                    return JsonResponse({"result": "ok"})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

            if action == 'edit':
                try:
                    oferta = OfertaLaboral.objects.get(pk=request.POST['id'])
                    f = OfertaLaboralForm(request.POST)
                    if f.is_valid():
                        if f.cleaned_data['fin'] < f.cleaned_data['inicio']:
                            return JsonResponse({"result": "bad", "mensaje": u"Fechas incorrectas."})
                        oferta.inicio = f.cleaned_data['inicio']
                        oferta.fin = f.cleaned_data['fin']
                        oferta.cargo = f.cleaned_data['cargo']
                        oferta.descripcion = f.cleaned_data['descripcion']
                        oferta.area = f.cleaned_data['area']
                        oferta.tiempo = f.cleaned_data['tiempo']
                        oferta.salario = f.cleaned_data['salario']
                        oferta.lugar = f.cleaned_data['lugar']
                        oferta.sexo = f.cleaned_data['sexo']
                        oferta.canton = f.cleaned_data['canton']
                        oferta.plazas = f.cleaned_data['plazas']
                        oferta.carreras = f.cleaned_data['carreras']
                        oferta.empresa = f.cleaned_data['empleador']
                        oferta.graduado = f.cleaned_data['graduado']
                        oferta.visibleinscrito = f.cleaned_data['visibleinscrito']
                        oferta.save(request)
                        log(u'Modifico oferta laboral: %s' % oferta, request, "edit")
                        return JsonResponse({"result": "ok", "id": oferta.id})
                    else:
                         raise NameError('Error')
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

            if action == 'observacion':
                try:
                    aplicanteoferta = AplicanteOferta.objects.get(pk=request.POST['id'])
                    f = ObservacionesOfertaLaboralForm(request.POST)
                    if f.is_valid():
                        aplicanteofertaobservacion = AplicanteOfertaObservacion(aplicanteoferta = aplicanteoferta,
                                                                                observacion = f.cleaned_data['observacion'])
                        aplicanteofertaobservacion.save(request)
                        aplicanteoferta.estado=False
                        aplicanteoferta.save(request)
                        log(u'Comentario - Observaciones: %s' % aplicanteoferta, request, "add")
                        send_html_mail("Comentario - Observaciones", "emails/observacion.html", {'sistema': request.session['nombresistema'],'observacion': aplicanteofertaobservacion.observacion,'oferta':aplicanteofertaobservacion.aplicanteoferta.oferta.cargo , 't': miinstitucion()}, aplicanteofertaobservacion.aplicanteoferta.inscripcion.persona.lista_emails(), [],  cuenta=CUENTAS_CORREOS[17][1])
                        return JsonResponse({"result": "ok"})
                    else:
                         raise NameError('Error')
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

            if action == 'editarea':
                try:
                    area = AreaOfertaLaboral.objects.get(pk=request.POST['id'])
                    f = AreaForm(request.POST)
                    if f.is_valid():
                        area.descripcion = f.cleaned_data['descripcion']
                        area.save(request)
                        log(u'Modifico area oferta laboral: %s' % area, request, "edit")
                        return JsonResponse({"result": "ok"})
                    else:
                         raise NameError('Error')
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

            if action == 'editpersonaresponsable':
                try:
                    personaresponsable = ConfiguraOferta.objects.get(pk=request.POST['id'])
                    f = ConfiguracionForm(request.POST)
                    if f.is_valid():
                        personaresponsable.administrativo_id = f.cleaned_data['administrativo']
                        personaresponsable.save(request)
                        log(u'Modifico persona responsable: %s' % personaresponsable, request, "edit")
                        return JsonResponse({"result": "ok"})
                    else:
                         raise NameError('Error')
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

            if action == 'add':
                try:
                    f = OfertaLaboralForm(request.POST)
                    if f.is_valid():
                        if f.cleaned_data['fin'] < datetime.now().date() or f.cleaned_data['fin'] < f.cleaned_data['inicio']:
                            return JsonResponse({"result": "bad", "mensaje": u"Fechas incorrectas."})
                        oferta = OfertaLaboral(empresa=f.cleaned_data['empleador'],
                                               registro=datetime.now().date(),
                                               inicio=f.cleaned_data['inicio'],
                                               fin=f.cleaned_data['fin'],
                                               cerrada=False,
                                               cargo=f.cleaned_data['cargo'],
                                               descripcion=f.cleaned_data['descripcion'],
                                               area=f.cleaned_data['area'],
                                               tiempo=f.cleaned_data['tiempo'],
                                               salario=f.cleaned_data['salario'],
                                               lugar=f.cleaned_data['lugar'],
                                               sexo=f.cleaned_data['sexo'],
                                               canton=f.cleaned_data['canton'],
                                               plazas=f.cleaned_data['plazas'],
                                               visibleinscrito=f.cleaned_data['visibleinscrito'],
                                               graduado=f.cleaned_data['graduado'],
                                               aprobado=True)
                        oferta.save(request)
                        oferta.carreras = f.cleaned_data['carreras']
                        log(u'Adiciono oferta laboral: %s' % oferta, request, "add")
                        return JsonResponse({"result": "ok", "id": oferta.id})
                    else:
                         raise NameError('Error')
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

            if action == 'activarinactivar':
                try:
                    area = AreaOfertaLaboral.objects.get(pk=request.POST['idarea'])
                    return JsonResponse({"result": "ok", "idarea": area.id,"descripcion": area.descripcion})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": "Error al guardar los datos."})

            if action == 'activarinactivarresponsable':
                try:
                    personaresponable = ConfiguraOferta.objects.get(pk=request.POST['idpers'])
                    return JsonResponse({"result": "ok", "idpersonaresponsable": personaresponable.id,"descripcion": personaresponable.administrativo.persona.apellido1 + ' ' + personaresponable.administrativo.persona.apellido2 + ' ' + personaresponable.administrativo.persona.nombres })
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": "Error al guardar los datos."})

            if action == 'aprobarestado':
                try:
                    area = AreaOfertaLaboral.objects.get(pk=request.POST['codigoitem'], status=True)
                    codigoactivainactiva = request.POST['codigoactivainactiva']
                    if codigoactivainactiva == '1':
                        area.estado = True
                    else:
                        area.estado = False
                    area.save(request)
                    log(u'Modificó estado: %s %s' % (area,area.estado), request, "edit")
                    return JsonResponse({"result": "ok"})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

            if action == 'aprobarestadoresponsable':
                try:
                    personaresponable = ConfiguraOferta.objects.get(pk=request.POST['codigoitem'], status=True)
                    codigoactivainactiva = request.POST['codigoactivainactiva']
                    if codigoactivainactiva == '1':
                        personaresponable.estado = True
                    else:
                        personaresponable.estado = False
                    personaresponable.save(request)
                    log(u'Modificó estado: %s %s' % (personaresponable,personaresponable.estado), request, "edit")
                    return JsonResponse({"result": "ok"})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

            if action == 'addarea':
                try:
                    f = AreaForm(request.POST)
                    if f.is_valid():
                        if AreaOfertaLaboral.objects.filter(descripcion=f.cleaned_data['descripcion'], status=True):
                            return JsonResponse({"result": "bad", "mensaje": u"Area ya existe."})
                        area = AreaOfertaLaboral(descripcion=f.cleaned_data['descripcion'])
                        area.save(request)
                        log(u'Adiciono area laboral: %s' % area, request, "add")
                        return JsonResponse({"result": "ok"})
                    else:
                         raise NameError('Error')
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

            if action == 'addpersonaresponsable':
                try:
                    f = ConfiguracionForm(request.POST)
                    if f.is_valid():
                        if ConfiguraOferta.objects.filter(administrativo_id=f.cleaned_data['administrativo'], status=True):
                            return JsonResponse({"result": "bad", "mensaje": u"Persona responsable ya existe."})
                        personaresponsable = ConfiguraOferta(administrativo_id=f.cleaned_data['administrativo'])
                        personaresponsable.save(request)
                        log(u'Adiciono persona responsable: %s' % personaresponsable, request, "add")
                        return JsonResponse({"result": "ok"})
                    else:
                         raise NameError('Error')
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

            elif action == 'aprobar':
                try:
                    aplicanteoferta = AplicanteOferta.objects.get(pk=request.POST['id'])
                    log(u'Aprobo solicitud oferta: %s' % aplicanteoferta, request, "edit")
                    aplicanteoferta.aprobada = True
                    aplicanteoferta.save(request)
                    return JsonResponse({"result": "ok"})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": u"Error al aprobar los datos."})

            elif action == 'validar':
                try:
                    aplicanteoferta = AplicanteOferta.objects.get(pk=request.POST['id'])
                    log(u'Validar solicitud oferta: %s' % aplicanteoferta, request, "edit")
                    aplicanteoferta.validada = True
                    aplicanteoferta.personavalidada = persona
                    aplicanteoferta.fechavalidada = datetime.now()
                    aplicanteoferta.save(request)
                    return JsonResponse({"result": "ok"})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": u"Error al validar los datos."})

            elif action == 'quitarvalidar':
                try:
                    aplicanteoferta = AplicanteOferta.objects.get(pk=request.POST['id'])
                    log(u'Quitar Validar solicitud oferta: %s' % aplicanteoferta, request, "edit")
                    aplicanteoferta.validada = False
                    aplicanteoferta.personavalidada = None
                    aplicanteoferta.fechavalidada = None
                    aplicanteoferta.save(request)
                    return JsonResponse({"result": "ok"})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": u"Error al validar los datos."})

            elif action == 'aprobaroferta':
                try:
                    ofertalaboral = OfertaLaboral.objects.get(pk=request.POST['id'])

                    log(u'Aprobo oferta Laboral: %s' % ofertalaboral, request, "edit")
                    ofertalaboral.aprobado = True
                    ofertalaboral.save(request)
                    return JsonResponse({"result": "ok"})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": u"Error al aprobar oferta."})
            elif action == 'gestionar':
                try:
                    ofertalaboral = OfertaLaboralEmpresa.objects.get(pk=int(encrypt(request.POST['id'])))
                    ofertalaboral.estadooferta = int(request.POST['estado'])
                    if int(request.POST['estado']) == 1:
                        estadooferta = 'aprobada'
                    elif int(request.POST['estado']) == 2:
                        estadooferta = 'rechazada'
                    else:
                        estadooferta = 'enviada a corregir'
                    if int(request.POST['estado']) >= 2 and 'motivo' in request.POST:
                        ofertalaboral.observacion_rechazo = request.POST['motivo']
                    else:
                        if variable_valor('NOTIFICA_EMPLEO'):
                            ids = []
                            if ofertalaboral.quienpostula == 0:
                                ins = Graduado.objects.filter(status=True, inscripcion__carrera_id__in=ofertalaboral.oferta_carreras_total()).values_list('inscripcion__persona_id', flat=True)
                                for p in ins:
                                    ids.append(p)
                            if ofertalaboral.quienpostula >= 1:
                                if ofertalaboral.quienpostula == 1:
                                    alum = Inscripcion.objects.filter(status=True, carrera_id__in=ofertalaboral.oferta_carreras_total(), coordinacion__id__lte=5).exclude(id__in=Graduado.objects.filter(status=True, inscripcion__carrera_id__in=ofertalaboral.oferta_carreras_total(), inscripcion__coordinacion__id__lte=5).values_list('inscripcion_id', flat=True))
                                else:
                                    alum = Inscripcion.objects.filter(status=True, carrera_id__in=ofertalaboral.oferta_carreras_total(), coordinacion__id__lte=5)
                                for alumno in alum:
                                    if alumno.mi_nivel().nivel_id >= 7:
                                        ids.append(alumno.persona_id)

                            personas = ids
                            nomempresa=ofertalaboral.encargado.empresa.nombre
                            nomoferta=ofertalaboral.titulo
                            if ofertalaboral.sexo > 0:
                                personas = Persona.objects.filter(status=True, id__in=ids, sexo_id=ofertalaboral.sexo).values_list('id', flat=True)
                            for p in personas:
                                # notificacion('Nueva oferta laboral', 'Existe una oferta laboral que te puede interesar', p, None,
                        #                             #              '/emp_postular', ofertalaboral.pk, 1, 'empleo',
                        #                             #              OfertaLaboralEmpresa, request)
                                notificacion = Notificacion(titulo='{}'.format(nomempresa),
                                                            cuerpo='Existe una nueva oferta laboral que te puede interesar: {}'.format(nomoferta),
                                                            destinatario_id=p,
                                                            # departamento=departamento,
                                                            url='/emp_postular?ofertaids={}'.format(encrypt(ofertalaboral.pk)),
                                                            object_id=ofertalaboral.pk,
                                                            prioridad=1,
                                                            app_label='empleo',
                                                            fecha_hora_visible=datetime.now() + timedelta(days=1)
                                                            )
                                notificacion.save(request)
                        else:
                            notificacion = Notificacion(titulo='Nueva oferta laboral',
                                                        cuerpo='Existe una oferta laboral que te puede interesar',
                                                        destinatario_id=22213,
                                                        # departamento=departamento,
                                                        url='/emp_postular',
                                                        object_id=ofertalaboral.pk,
                                                        prioridad=1,
                                                        app_label='empleo',
                                                        fecha_hora_visible=datetime.now() + timedelta(days=1)
                                                        )
                            notificacion.save(request)


                    notificacionempresa = Notificacion(titulo='Oferta laboral {}'.format(estadooferta),
                                                       cuerpo='La oferta laboral "<b>{}</b>" fue {}'.format(
                                                           ofertalaboral.titulo, estadooferta),
                                                       destinatario_id=ofertalaboral.empresa.persona.id,
                                                       url='/empresa/empr_ofertas?id={}'.format(request.POST['id']),
                                                       object_id=ofertalaboral.pk,
                                                       prioridad=1,
                                                       app_label='empresa',
                                                       fecha_hora_visible=datetime.now() + timedelta(days=1)
                                                       )
                    notificacionempresa.save(request)
                    ofertalaboral.save(request)
                    log(u'Gestiono oferta Laboral: %s' % ofertalaboral, request, "edit")
                    send_html_mail('Estado de la oferta: {}'.format(estadooferta),
                                   "emails/estadoofertalaboral.html",
                                   {'sistema': u'Unemi Empleo', 'fecha': datetime.now(),
                                    'ofertalaboral': ofertalaboral, 'tit': 'Unemi - Empleo', 'estadooferta': estadooferta},
                                   ofertalaboral.encargado.persona.lista_emails_envio(),
                                   [], cuenta=CUENTAS_CORREOS[17][1])

                    return JsonResponse({"result": "ok"})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": u"Error al aprobar oferta."})
            elif action == 'gestionarempresa':
                try:
                    empresa = SolicitudAprobacionEmpresa.objects.get(pk=int(encrypt(request.POST['id'])))
                    empresa.estadoempresa = int(request.POST['estado'])

                    if int(request.POST['estado']) >= 2:
                        estadoempresa = 'rechazado'
                    else:
                        estadoempresa = 'aprobado'

                    if int(request.POST['estado']) >= 2:
                        empresa.observacion = request.POST['motivo']

                    empresa.save(request)
                    log(u'Gestiono solicitud empresa: %s' % empresa, request, "edit")

                    historial = HistorialAprobacionEmpresa(empleador=empresa.empleador, solicitud=empresa, fecha=datetime.now())
                    historial.save(request)
                    datospersona = empresa.empleador.persona
                    empleador = empresa.empleador
                    if datospersona:
                        username_ = None
                        pass_ = None
                        if not datospersona.usuario:
                            username_ = calculate_username_empresa(empresa.empleador)
                            generar_usuario_cedula(datospersona, username_, 356)
                            # password, anio = '', ''
                            # if datospersona.nacimiento:
                            #     anio = "*" + str(datospersona.nacimiento)[0:4]
                            password = datospersona.ruc.strip()
                            mensaje = 'Su usuario es: {} y su clave temporal es su número de ruc: {}'.format(username_,
                                                                                                             password)
                        else:
                            user_ = User.objects.get(pk=datospersona.usuario.pk)
                            username_ = user_.username
                            group_ = Group.objects.get(pk=392)
                            group_.user_set.add(user_)
                            group_.save()
                            existia = True
                            mensaje = 'Registro exitoso, acceda con sus credenciales del Sistema de Gestión Academica (SGA)'
                    perfil = PerfilUsuario(persona=datospersona, empleador=empresa.empleador)
                    perfil.save(request)
                    empleador.autorizada = True
                    empleador.save(request)
                    send_html_mail(u'Estado de registro de empresa: {}'.format(estadoempresa), "emails/registroempresa.html",
                                   {'sistema': u'Unemi Empresas', 'fecha': datetime.now(),
                                    'empresa': empresa, 'tit': 'Unemi - Empresa'},
                                   empresa.empleador.persona.lista_emails_envio(), [], cuenta=CUENTAS_CORREOS[17][1])
                    return JsonResponse({"result": "ok"})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": u"Error al aprobar oferta."})

            elif action == 'deleteempresa':
                try:
                    empresa = Empleador.objects.get(pk=int(encrypt(request.POST['id'])))
                    empresa.status = False
                    empresa.save(request)
                    log(u'Elimino empresa: %s' % empresa, request, "delete")

                    return JsonResponse({"result": "ok"})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar empresa."})

            elif action == 'deletesolicitud':
                try:
                    solicitud = SolicitudRevisionTitulo.objects.filter(status=True, id=int(encrypt(request.POST['id']))).first()
                    if not solicitud:
                        raise NameError('Solicitud no encontrada')
                    solicitud.status = False
                    solicitud.save(request, update_fields=['status'])
                    log('Eliminó solicitud de revisión de titulo: {}'.format(
                        solicitud.inscripcion.persona.nombre_completo_minus()), request, "delete")
                    return JsonResponse({"error": False})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": True, "mensaje": "{}".format(ex)})

            elif action == 'gestionarsolicitud':
                try:
                    from empleo.forms import GestionarSolicitudRevisionTituloForm
                    form = GestionarSolicitudRevisionTituloForm(request.POST)
                    if form.is_valid():
                        solicitud = SolicitudRevisionTitulo.objects.filter(id=request.POST['id']).first()
                        if not solicitud:
                            raise NameError('Solicitud no encontrada')
                        solicitud.estado = form.cleaned_data['estado']
                        solicitud.observacion = form.cleaned_data['observacion']
                        solicitud.fechaaprueba = datetime.now().date()
                        solicitud.personaaprueba = persona
                        solicitud.save(request, update_fields=['estado', 'observacion', 'fechaaprueba', 'personaaprueba'])
                        log('Gestionó solicitud de revisión de titulo: {}'.format(solicitud.inscripcion.persona.nombre_completo_minus()), request, "edit")
                        # notificacion('Solicitud de revisión de titulo', 'El estudiante {} ingresó una solicitud de revisión de titulo'.format(solicitud.inscripcion.persona.nombre_completo_minus()), para, None, '/pro_cronograma?action=listatutorias', add.pk, 1,
                        #              'sga', InquietudPracticasPreprofesionales, request)
                        return JsonResponse({'result': False})
                except Exception as ex:
                    pass
        return JsonResponse({"result": "bad", "mensaje": u"Solicitud Incorrecta."})
    else:
        if 'action' in request.GET:
            data = {}
            data['action'] = action = request.GET['action']
            adduserdata(request, data)

            if action == 'verhojavida':
                try:
                    data['title'] = u"Hoja de vida del aplicante"
                    data['aplicante'] = AplicanteOferta.objects.get(pk=request.GET['id'])
                    data['experiencias'] = data['aplicante'].inscripcion.tiene_experiencia()
                    data['estudios'] = data['aplicante'].inscripcion.tiene_estudios()
                    data['idiomas'] = data['aplicante'].inscripcion.persona.tiene_idiomadomina()
                    data['conocimientos'] = data['aplicante'].inscripcion.tiene_conocimientos()
                    data['conocimientosadicionales'] = data['aplicante'].inscripcion.persona.tiene_conocimientosadicionales()
                    data['referencias'] = data['aplicante'].inscripcion.persona.tiene_referencias()
                    data['nivel'] = NIVEL_CONOCIMIENTO
                    if ConocimientoInformatico.objects.filter(inscripcion=data['aplicante'].inscripcion).exists():
                        co = []
                        conocimientocategoria = ConocimientoInformatico.objects.filter(inscripcion=data['aplicante'].inscripcion)
                        for cono in conocimientocategoria:
                            if cono.herramienta.categoria.id not in co:
                                co.append(cono.herramienta.categoria.id)
                        data['categoriaherramienta'] = CategoriaHerramienta.objects.filter(id__in=co)
                    else:
                        data['categoriaherramienta'] = []
                    return render(request, "adm_ofertalaboral/verhojavida.html", data)
                except Exception as ex:
                    pass

            elif action == 'cerrar':
                try:
                    oferta = OfertaLaboral.objects.get(pk=request.GET['id'])
                    oferta.cerrada = True
                    oferta.save(request)
                    log(u'Cerro Oferta laboral: %s' % oferta, request, "edit")
                    return HttpResponseRedirect("/adm_ofertalaboral?id=" + request.GET['id'])
                except Exception as ex:
                    pass

            elif action == 'abrir':
                try:
                    oferta = OfertaLaboral.objects.get(pk=request.GET['id'])
                    oferta.cerrada = False
                    if not oferta.fin >= datetime.now().date():
                        oferta.fin = datetime.now().date()
                    oferta.save(request)
                    log(u'Abrio Oferta laboral: %s' % oferta, request, "edit")
                    return HttpResponseRedirect("/adm_ofertalaboral?id=" + request.GET['id'])
                except Exception as ex:
                    pass

            elif action == 'aprobaroferta':
                try:
                    data['title'] = u'Aprobar oferta laboral'
                    data['ofertalaboral'] = OfertaLaboral.objects.get(pk=request.GET['id'])
                    return render(request, "adm_ofertalaboral/aprobaroferta.html", data)
                except Exception as ex:
                    pass

            elif action == 'asigcita':
                try:
                    data['title'] = u"Asignar cita"
                    registrado = AplicanteOferta.objects.get(pk=request.GET['id'])
                    form = AgendarCitaForm(initial={"horaentrevista": registrado.horaentrevista.__str__() if registrado.horaentrevista else "12:00",
                                                    "fechaentrevista": registrado.fechaentrevista if registrado.fechaentrevista else datetime.now().date(),
                                                    "lugar": registrado.lugar if registrado.lugar else registrado.oferta.empresa.direccion,
                                                    "personacontacto": registrado.personacontacto if registrado.personacontacto else data['persona'],
                                                    "telefonocontacto": registrado.telefonocontacto if registrado.telefonocontacto else registrado.oferta.empresa.telefonos})
                    data['registrado'] = registrado
                    data['form'] = form
                    return render(request, "adm_ofertalaboral/asigcita.html", data)
                except Exception as ex:
                    pass

            elif action == 'registrados':
                try:
                    data['title'] = u"Lista de registrados"
                    data['idoferta'] = request.GET['id']
                    oferta = OfertaLaboral.objects.get(pk=request.GET['id'])
                    data['registrados'] = oferta.aplicanteoferta_set.all().order_by('inscripcion__persona__nombres')
                    data['reporte_0'] = obtener_reporte('hoja_vida_sagest')
                    data['reporte_1'] = obtener_reporte('registro_postulantes')
                    return render(request, "adm_ofertalaboral/registrados.html", data)
                except Exception as ex:
                    pass

            elif action == "del":
                try:
                    data['title'] = u"Eliminar oferta laboral"
                    oferta = OfertaLaboral.objects.get(pk=request.GET['id'])
                    data['oferta'] = oferta
                    return render(request, "adm_ofertalaboral/del.html", data)
                except Exception as ex:
                    pass

            elif action == "delarea":
                try:
                    data['title'] = u"Eliminar area oferta laboral"
                    area = AreaOfertaLaboral.objects.get(pk=request.GET['id'])
                    data['area'] = area
                    return render(request, "adm_ofertalaboral/delarea.html", data)
                except Exception as ex:
                    pass

            elif action == "delpersonaresponsable":
                try:
                    data['title'] = u"Eliminar persona responsable"
                    personaresponable = ConfiguraOferta.objects.get(pk=request.GET['id'])
                    data['personaresponable'] = personaresponable
                    return render(request, "adm_ofertalaboral/delpersonaresponsable.html", data)
                except Exception as ex:
                    pass

            elif action == "add":
                try:
                    data['title'] = u"Nueva oferta laboral"
                    form = OfertaLaboralForm(initial={"inicio": datetime.now().date(),
                                                      "fin": (datetime.now() + timedelta(days=30)).date()})
                    data['form'] = form
                    return render(request, "adm_ofertalaboral/add.html", data)
                except Exception as ex:
                    pass
#------------------------------------------------------------------------------------------------------------
            elif action == "editgrupocar":
                try:
                    data['title'] = u"Editar Grupo Carrera"
                    grupocarrera = CarreraGrupo.objects.get(pk=encrypt(request.GET['id']))
                    form = GrupoCarreraForm(initial={"nombre": grupocarrera.nombre, "activa": grupocarrera.activa})
                    data['form'] = form
                    data['grupocarrera'] = grupocarrera

                    #return render(request, "empresa/editargrupocarrera.html", data)
                    template = get_template("empresa/editargrupocarrera.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass
#--------------------------------------------------------------------------------------------------
            elif action == "agregargrupocarrera":
                try:
                    data['title'] = u"Nuevo Grupo Carrera"
                    form = GrupoCarreraForm()
                    data['form'] = form
                    template = get_template("empresa/addgruposcarrera.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass
#--------------------------------------------------------------------------------------------

            elif action == "observacion":
                try:
                    data['title'] = u"Comentarios - Obserbaciones"
                    aplicanteoferta = AplicanteOferta.objects.get(pk=request.GET['id'])
                    form = ObservacionesOfertaLaboralForm()
                    data['form'] = form
                    data['aplicanteoferta'] = aplicanteoferta
                    return render(request, "adm_ofertalaboral/observacion.html", data)
                except Exception as ex:
                    pass

            elif action == "edit":
                try:
                    data['title'] = u"Editar oferta laboral"
                    oferta = OfertaLaboral.objects.get(pk=request.GET['id'])
                    form = OfertaLaboralForm(initial={"inicio": oferta.inicio,
                                                      "fin": oferta.fin,
                                                      "cargo": oferta.cargo,
                                                      "descripcion": oferta.descripcion,
                                                      "area": oferta.area,
                                                      "graduado": oferta.graduado,
                                                      "tiempo": oferta.tiempo,
                                                      "salario": oferta.salario,
                                                      "lugar": oferta.lugar,
                                                      "sexo": oferta.sexo,
                                                      "canton": oferta.canton,
                                                      "carreras": oferta.carreras.all(),
                                                      "plazas": oferta.plazas,
                                                      "visibleinscrito": oferta.visibleinscrito,
                                                      "empleador": oferta.empresa})
                    data['form'] = form
                    data['oferta'] = oferta
                    return render(request, "adm_ofertalaboral/edit.html", data)
                except Exception as ex:
                    pass

            elif action == "editarea":
                try:
                    data['title'] = u"Editar area laboral"
                    area = AreaOfertaLaboral.objects.get(pk=request.GET['id'])
                    form = AreaForm(initial={"descripcion": area.descripcion})
                    data['form'] = form
                    data['area'] = area
                    return render(request, "adm_ofertalaboral/editarea.html", data)
                except Exception as ex:
                    pass

            elif action == "editpersonaresponsable":
                try:
                    data['title'] = u"Editar persona responsable"
                    personaresponable = ConfiguraOferta.objects.get(pk=request.GET['id'])
                    form = ConfiguracionForm(initial={"administrativo": personaresponable.administrativo.id })
                    data['form'] = form
                    data['personaresponable'] = personaresponable
                    return render(request, "adm_ofertalaboral/editpersonaresponsable.html", data)
                except Exception as ex:
                    pass


            elif action == 'addarea':
                try:
                    data['title'] = u"Nueva Area"
                    form = AreaForm()
                    data['form'] = form
                    return render(request, "adm_ofertalaboral/addarea.html", data)
                except Exception as ex:
                    pass

            elif action == 'addpersonaresponsable':
                try:
                    data['title'] = u"Nueva Persona Responsable"
                    form = ConfiguracionForm()
                    data['form'] = form
                    return render(request, "adm_ofertalaboral/addpersonaresponsable.html", data)
                except Exception as ex:
                    pass

            elif action == 'verareas':
                try:
                    data['title'] = u'Listado de areas registradas'
                    search = None
                    ids = None
                    if 'se' in request.GET:
                        search = request.GET['se']
                        area = AreaOfertaLaboral.objects.filter(descripcion__icontains=search, status=True)
                    elif 'ide' in request.GET:
                        ids = request.GET['ide']
                        area = AreaOfertaLaboral.objects.filter(id=ids,status=True)
                    else:
                        area = AreaOfertaLaboral.objects.filter(status=True).order_by('descripcion')
                    paging = MiPaginador(area, 25)
                    p = 1
                    try:
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        page = paging.page(p)
                    except Exception as ex:
                        page = paging.page(p)
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['areas'] = page.object_list
                    return render(request, "adm_ofertalaboral/verareas.html", data)
                except Exception as ex:
                    pass

            elif action == 'verconfiguracion':
                try:
                    data['title'] = u'Listado de personas responsable'
                    search = None
                    ids = None
                    if 'se' in request.GET:
                        search = request.GET['se']
                        ss = search.split(' ')
                        if len(ss) == 1:
                            personaresponsable = ConfiguraOferta.objects.filter(Q(administrativo__persona__nombres__icontains=search) |
                                                                                Q(administrativo__persona__apellido1__icontains=search) |
                                                                                Q(administrativo__persona__apellido2__icontains=search),
                                                                                status=True)
                        else:
                            personaresponsable = ConfiguraOferta.objects.filter(Q(administrativo__persona__apellido1__icontains=ss[0]) &
                                                                                Q(administrativo__persona__apellido2__icontains=ss[1]), status=True)
                    elif 'ide' in request.GET:
                        ids = request.GET['ide']
                        personaresponsable = ConfiguraOferta.objects.filter(id=ids,status=True)
                    else:
                        personaresponsable = ConfiguraOferta.objects.filter(status=True).order_by('administrativo')
                    paging = MiPaginador(personaresponsable, 25)
                    p = 1
                    try:
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        page = paging.page(p)
                    except Exception as ex:
                        page = paging.page(p)
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['personaresponsables'] = page.object_list
                    return render(request, "adm_ofertalaboral/verconfiguracion.html", data)
                except Exception as ex:
                    pass

            elif action == 'detalle':
                try:
                    data = {}
                    data['ofertalaboral'] = ofertalaboral = OfertaLaboral.objects.get(pk=int(request.GET['id']))
                    data['carreras'] = ofertalaboral.vercarreras()
                    template = get_template("adm_ofertalaboral/detalle.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

            elif action == 'aprobar':
                try:
                    data['title'] = u'Aprobar Solicitud'
                    data['aplicanteoferta'] = AplicanteOferta.objects.get(pk=request.GET['id'])
                    return render(request, "adm_ofertalaboral/aprobar.html", data)
                except Exception as ex:
                    pass

            elif action == 'validar':
                try:
                    data['title'] = u'Validar Solicitud'
                    data['aplicanteoferta'] = aplicanteoferta = AplicanteOferta.objects.get(pk=request.GET['id'])
                    if not aplicanteoferta.validada:
                        return render(request, "adm_ofertalaboral/validar.html", data)
                    else:
                        return render(request, "adm_ofertalaboral/quitarvalidar.html", data)
                except Exception as ex:
                    pass

            # aprobacion de empreas de unemi empleo
            elif action == 'aprobarempresa':
                try:
                    id = request.GET['id']
                    solicitud = SolicitudAprobacionEmpresa.objects.get(id=id)
                    # log(u'Gestiono empresa: %s' % solicitud.empleador, request, "edit")
                    # solicitud.estado = request.POST['estado']
                    # if int(request.POST['estado']) >= 2:
                    #     solicitud.observacion = request.POST['observacion_rechazo']
                    # else:
                    datospersona = solicitud.empleador.persona
                    empleador = solicitud.empleador
                    if datospersona:
                        username_ = None
                        pass_ = None
                        if not datospersona.usuario:
                            username_ = calculate_username_empresa(solicitud.empleador)
                            generar_usuario_cedula(datospersona, username_, 356)
                            # password, anio = '', ''
                            # if datospersona.nacimiento:
                            #     anio = "*" + str(datospersona.nacimiento)[0:4]
                            password = datospersona.ruc.strip()
                            mensaje = 'Su usuario es: {} y su clave temporal es su número de ruc: {}'.format(username_,
                                                                                                             password)
                        else:
                            user_ = User.objects.get(pk=datospersona.usuario.pk)
                            username_ = user_.username
                            group_ = Group.objects.get(pk=392)
                            group_.user_set.add(user_)
                            group_.save()
                            existia = True
                            mensaje = 'Registro exitoso, acceda con sus credenciales del Sistema de Gestión Academica (SGA)'
                    perfil = PerfilUsuario(persona=datospersona, empleador=solicitud.empleador)
                    perfil.save(request)
                    solicitud.estadoempresa = 1
                    solicitud.save()
                    empleador.autorizada = True
                    empleador.save(request)
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos. %s" % ex})

            elif action == 'verofertas':
                try:
                    adduserdata(request, data)
                    empleador = Empleador.objects.get(id=request.GET['id'])
                    data['id'] = id = request.GET['id']
                    data['title'] = 'Listado de ofertas laborales'
                    data['fechaactual'] = fechaactual = datetime.now().date()
                    search = None
                    ids = None
                    url_vars = ''
                    filtro = (Q(status=True))
                    if 's' in request.GET:
                        data['s'] = s = request.GET['s']
                        url_vars += "&s={}".format(s)
                        filtro = filtro & (Q(titulo__icontains=s))
                    ofertas = empleador.ofertalaboralempresa_set.filter(filtro)
                    paging = MiPaginador(ofertas, 25)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)

                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data['ofertas'] = page.object_list
                    data['empresa'] = empleador
                    data['total'] = len(ofertas)
                    if 'exportar_excel' in request.GET:
                        wb = openxl.Workbook()
                        wb["Sheet"].title = "Reporte_ofertas_postulantes"
                        ws = wb.active
                        style_title = openxlFont(name='Arial', size=16, bold=True)
                        style_cab = openxlFont(name='Arial', size=10, bold=True)
                        alinear = alin(horizontal="center", vertical="center")
                        alinearizq = alin(horizontal="left", vertical="center")
                        response = HttpResponse(content_type="application/ms-excel")
                        response[
                            'Content-Disposition'] = 'attachment; filename=Reporte de ofertas laborales con postulantes' + '-' + random.randint(
                            1, 10000).__str__() + '.xlsx'
                        ws.column_dimensions['B'].width = 35
                        ws.column_dimensions['C'].width = 35
                        ws.column_dimensions['D'].width = 15
                        ws.column_dimensions['E'].width = 15
                        ws.column_dimensions['F'].width = 35
                        ws.column_dimensions['G'].width = 20
                        ws.column_dimensions['H'].width = 20
                        ws.column_dimensions['I'].width = 20
                        ws.column_dimensions['J'].width = 20
                        ws.column_dimensions['K'].width = 20
                        ws.merge_cells('A1:K1')
                        ws['A1'] = 'OFERTAS LABORALES - POSTULANTES'
                        celda1 = ws['A1']
                        celda1.font = style_title
                        celda1.alignment = alinear

                        ws.merge_cells('A2:C2')
                        ws['A2'] = 'EMPRESA: ' + str(empleador.nombre)
                        celda1 = ws['A2']
                        celda1.font = style_cab
                        celda1.alignment = alinearizq

                        columns = [u"N°", u"OFERTA LABORAL", u"NOMBRES", u"CEDULA", u"TELEFONO", u"EMAIL",
                                   u"F. POSTULACIÓN", u"F. REVISIÓN", u"HOJA DE VIDA", u"CONTRATADO"
                                   ]
                        row_num = 3
                        for col_num in range(0, len(columns)):
                            celda = ws.cell(row=row_num, column=(col_num + 1), value=columns[col_num])
                            celda.font = style_cab
                        row_num = 4
                        numero = 1
                        for list in ofertas:
                            postulantes = list.participantes()
                            for post in postulantes:
                                ws.cell(row=row_num, column=1, value=numero)
                                ws.cell(row=row_num, column=2, value=str(list.titulo))
                                ws.cell(row=row_num, column=3, value=str(post.persona.nombre_completo_minus()))
                                ws.cell(row=row_num, column=4, value=str(post.persona.cedula))
                                ws.cell(row=row_num, column=5, value=str(post.persona.telefono))
                                ws.cell(row=row_num, column=6, value=str(
                                    post.persona.emailinst) if post.persona.emailinst else post.persona.email)
                                ws.cell(row=row_num, column=7, value=str(post.fecha_creacion.date()))
                                ws.cell(row=row_num, column=8,
                                        value=str(post.fecha_revision.date()) if post.fecha_revision else 'Sin revisar')
                                ws.cell(row=row_num, column=9, value=str(post.get_estado_display()))
                                if post.estado == 2:
                                    if post.estcontrato == 0:
                                        estadocontrato = "Pendiente"
                                    if post.estcontrato == 1:
                                        estadocontrato = "Si"
                                    if post.estcontrato == 2:
                                        estadocontrato = "No"
                                else:
                                    estadocontrato = "-"
                                ws.cell(row=row_num, column=10, value=str(estadocontrato))
                                row_num += 1
                                numero += 1
                        wb.save(response)
                        return response

                    return render(request, "adm_ofertalaboral/view_ofertas_unemi_empleo.html", data)

                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos. %s" % ex})
            elif action == 'verofertastotal':
                try:
                    adduserdata(request, data)
                    data['title'] = 'Listado total de ofertas laborales'
                    search = None
                    ids = None
                    url_vars = f'&action={action}'
                    filtro = (Q(status=True))
                    ofertas = OfertaLaboralEmpresa.objects.filter(status=True)
                    data['empresas'] = empresas = Empleador.objects.filter(status=True, id__in=ofertas.values_list('empresa_id', flat=True))
                    data['encargados'] = encargados = RepresentantesEmpresa.objects.filter(status=True, id__in=ofertas.values_list('encargado_id', flat=True))
                    empresaselect = 0
                    encargadoselect = 0
                    if 's' in request.GET:
                        data['s'] = s = request.GET['s']
                        url_vars += "&s={}".format(s)
                        filtro = filtro & (Q(empresa__persona__apellido1__icontains=s) | Q(empresa__persona__apellido2__icontains=s) | Q(
                            empresa__persona__nombres__icontains=s) | Q(titulo__icontains=s))
                    if 'empresa' in request.GET :
                        data['empresa'] = s = int(request.GET['empresa'])
                        if s > 0:
                            empresaselect = s
                            url_vars += "&empresa={}".format(s)
                            filtro = filtro & Q(empresa_id=s)
                    if 'encargado' in request.GET:
                        data['encargado'] = s = int(request.GET['encargado'])
                        if s > 0:
                            encargadoselect = s
                            url_vars += "&encargado={}".format(s)
                            filtro = filtro & Q(encargado_id=s)
                    ofertas = ofertas.filter(filtro)
                    paging = MiPaginador(ofertas, 25)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)

                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data['ofertas'] = page.object_list
                    data['empresaselect'] = empresaselect
                    data['encargadoselect'] = encargadoselect
                    data['total']=len(ofertas)
                    if 'exportar_excel' in request.GET:
                        wb = openxl.Workbook()
                        wb["Sheet"].title = "Reporte_ofertas_laborales"
                        ws = wb.active
                        style_title = openxlFont(name='Arial', size=16, bold=True)
                        style_cab = openxlFont(name='Arial', size=10, bold=True)
                        alinear = alin(horizontal="center", vertical="center")
                        response = HttpResponse(content_type="application/ms-excel")
                        response[
                            'Content-Disposition'] = 'attachment; filename=Reporte de ofertas laborales' + '-' + random.randint(
                            1, 10000).__str__() + '.xlsx'
                        ws.column_dimensions['B'].width = 25
                        ws.column_dimensions['C'].width = 25
                        ws.column_dimensions['D'].width = 15
                        ws.column_dimensions['E'].width = 25
                        ws.column_dimensions['F'].width = 25
                        ws.column_dimensions['G'].width = 25
                        ws.column_dimensions['H'].width = 25
                        ws.column_dimensions['I'].width = 25
                        ws.column_dimensions['J'].width = 25
                        ws.column_dimensions['K'].width = 25
                        ws.column_dimensions['L'].width = 25
                        ws.column_dimensions['M'].width = 25
                        ws.merge_cells('A1:M1')
                        ws['A1'] = 'OFERTAS LABORALES'
                        celda1 = ws['A1']
                        celda1.font = style_title
                        celda1.alignment = alinear

                        columns = [u"N°", u"EMPRESA", u"NOMBRE", u"ENCARGADO", u"NIVEL MÍNIMO",
                                   u"MODALIDAD",u"DEDICACIÓN",u"JORNADA",u"RMU",u"TIPO CONTRATO", u"EXPERIENCIA", u"VACANTES", u"VIGENCIA"
                                   ]
                        row_num = 3
                        for col_num in range(0, len(columns)):
                            celda = ws.cell(row=row_num, column=(col_num + 1), value=columns[col_num])
                            celda.font = style_cab
                        row_num = 4
                        numero = 1
                        for list in ofertas:
                            fechas=f"{list.finicio} a {list.ffin}"
                            ws.cell(row=row_num, column=1, value=numero)
                            ws.cell(row=row_num, column=2, value=str(list.empresa.persona.nombre_minus()))
                            ws.cell(row=row_num, column=3, value=str(list.titulo))
                            ws.cell(row=row_num, column=4, value=str(list.encargado.persona.nombre_completo_minus()))
                            ws.cell(row=row_num, column=5, value=str(list.get_nivel_display()))
                            ws.cell(row=row_num, column=6, value=str(list.get_modalidad_display()))
                            ws.cell(row=row_num, column=7, value=str(list.get_dedicacion_display()))
                            ws.cell(row=row_num, column=8, value=str(list.get_jornada_display()))
                            ws.cell(row=row_num, column=9, value=str(list.rmu))
                            ws.cell(row=row_num, column=10, value=str(list.tipocontrato).capitalize() if list.tipocontrato else "No registra")
                            ws.cell(row=row_num, column=11, value=str(list.get_tiempoexperiencia_display()).capitalize() if list.requiereexpe else "No requiere experiencia")
                            ws.cell(row=row_num, column=12, value=str(list.vacantes))
                            ws.cell(row=row_num, column=13, value=fechas)
                            row_num += 1
                            numero += 1
                        wb.save(response)
                        return response
                    return render(request, "adm_ofertalaboral/lista_total_ofertas.html", data)

                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos. %s" % ex})

            if action == 'verdetalle':
                try:
                    data['id'] = id = int(encrypt(request.GET['id']))
                    data['nopostula'] =True
                    data['filtro'] = filtro = OfertaLaboralEmpresa.objects.get(pk=id)
                    data['hoy'] = hoy = datetime.now().date()
                    template = get_template("empleo/postular/verdetalle.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass
            if action == 'gestionar':
                try:
                    data['id'] = id = int(encrypt(request.GET['id']))
                    data['filtro'] = filtro = OfertaLaboralEmpresa.objects.get(pk=id)
                    form= GestionarOfertaForm()
                    form.gestionar(filtro.estadooferta)
                    data['form'] = form
                    template = get_template("adm_ofertalaboral/gestionar.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass
            if action == 'gestionarempresa':
                try:
                    data['id'] = id = int(encrypt(request.GET['id']))
                    data['filtro'] = filtro = SolicitudAprobacionEmpresa.objects.get(id=id)
                    data['form'] = GestionarEmpresaForm()
                    template = get_template("adm_ofertalaboral/gestionarempresa.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass
            elif action == 'participantes':
                try:
                    data['title'] = 'Lista de postulantes'
                    data['id'] = id = int(encrypt(request.GET['id']))
                    data['oferta'] = oferta = OfertaLaboralEmpresa.objects.get(pk=id)
                    data['postulantes'] = postulantes = oferta.participantes()
                    data['empresa'] = oferta.empresa

                    estado, search, filtro, url_vars = request.GET.get('estado', ''),request.GET.get('search', ''), (Q(status=True)), f'&action={action}&id={request.GET["id"]}'
                    if estado:
                        data['estadoapto'] = int(estado)
                        url_vars += "&estado={}".format(estado)
                        postulantes = postulantes.filter(
                            Q(estado=estado))
                    if search:
                        data['search'] = search = request.GET['search'].strip()
                        url_vars += "&search={}".format(search)
                        postulantes = postulantes.filter(Q(persona__apellido1__icontains=search) | Q(persona__apellido2__icontains=search) | Q(persona__nombres__icontains=search) | Q(persona__cedula__icontains=search)).distinct()

                    paging = MiPaginador(postulantes, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['listado'] = page.object_list
                    data["url_vars"] = url_vars
                    data['list_count'] = len(postulantes)
                    data['estado'] = ESTADOS_APLICACION
                    if 'exportar_excel' in request.GET:
                        wb = openxl.Workbook()
                        wb["Sheet"].title = "Reporte_postulantes"
                        ws = wb.active
                        style_title = openxlFont(name='Arial', size=16, bold=True)
                        style_cab = openxlFont(name='Arial', size=10, bold=True)
                        alinear = alin(horizontal="center", vertical="center")
                        response = HttpResponse(content_type="application/ms-excel")
                        response[
                            'Content-Disposition'] = 'attachment; filename=Reporte de postulantes' + '-' + random.randint(
                            1, 10000).__str__() + '.xlsx'
                        ws.column_dimensions['B'].width = 30
                        ws.column_dimensions['C'].width = 20
                        ws.column_dimensions['D'].width = 15
                        ws.column_dimensions['E'].width = 25
                        ws.column_dimensions['F'].width = 20
                        ws.column_dimensions['G'].width = 20
                        ws.column_dimensions['H'].width = 20
                        ws.column_dimensions['I'].width = 20
                        ws.merge_cells('A1:I1')
                        ws['A1'] = 'POSTULANTES DE OFERTAS LABORALES'
                        celda1 = ws['A1']
                        celda1.font = style_title
                        celda1.alignment = alinear

                        ws.merge_cells('A2:B2')
                        ws['A2'] = 'Ofertal laboral: ' + str(oferta.titulo)
                        celda1 = ws['A2']
                        celda1.font = style_cab
                        celda1.alignment = alinear

                        columns = [u"N°", u"NOMBRE", u"CEDULA", u"TELEFONO", u"EMAIL",
                                   u"F. POSTULACIÓN", u"F. REVISIÓN", u"HOJA DE VIDA", u"CONTRATADO",
                                   ]
                        row_num = 3
                        for col_num in range(0, len(columns)):
                            celda = ws.cell(row=row_num, column=(col_num + 1), value=columns[col_num])
                            celda.font = style_cab
                        row_num = 4
                        numero = 1
                        estadocontrato = ""
                        for list in postulantes:
                            ws.cell(row=row_num, column=1, value=numero)
                            ws.cell(row=row_num, column=2, value=str(list.persona.nombre_completo_minus()))
                            ws.cell(row=row_num, column=3, value=str(list.persona.cedula))
                            ws.cell(row=row_num, column=4, value=str(list.persona.telefono))
                            ws.cell(row=row_num, column=5, value=str(list.persona.emailinst) if list.persona.emailinst else list.persona.email)
                            ws.cell(row=row_num, column=6, value=str(list.fecha_creacion.date()))
                            ws.cell(row=row_num, column=7, value=str(list.fecha_revision.date()) if list.fecha_revision else 'Sin revisar')
                            ws.cell(row=row_num, column=8, value=str(list.get_estado_display()))
                            if list.estado == 2:
                                if list.estcontrato ==0:
                                    estadocontrato="Pendiente"
                                if list.estcontrato == 1:
                                    estadocontrato = "Si"
                                if list.estcontrato == 2:
                                    estadocontrato = "No"
                            else:
                                estadocontrato = "-"
                            ws.cell(row=row_num, column=9, value=str(estadocontrato))
                            row_num += 1
                            numero += 1
                        wb.save(response)
                        return response

                    return render(request, "adm_ofertalaboral/participantes.html", data)
                except Exception as ex:
                    pass

            if action == 'editsbu':
                try:
                    data['title'] = u'Modificar el Salario'
                    data['action'] = request.GET['action']
                    data['sbu_a'] = variable_valor('SBU_VALOR')
                    template = get_template("adm_ofertalaboral/modal/formularioSBU.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, 'message': str(ex)})

            elif action == 'participantestotal':
                try:
                    data['title'] = 'Lista de postulantes'
                    data['postulantes'] = postulantes = PersonaAplicaOferta.objects.filter(status=True)

                    search, filtro, url_vars = request.GET.get('search', ''), (Q(status=True)), ''
                    if search:
                        data['search'] = search = request.GET['search'].strip()
                        url_vars += "&search={}".format(search)
                        postulantes = postulantes.filter(Q(persona__apellido1__icontains=search) | Q(persona__apellido2__icontains=search) | Q(persona__nombres__icontains=search) | Q(persona__cedula__icontains=search)).distinct()

                    paging = MiPaginador(postulantes, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['listado'] = page.object_list
                    data["url_vars"] = url_vars
                    data['list_count'] = len(postulantes)
                    return render(request, "empresa/participanteslist.html", data)
                except Exception as ex:
                    pass
#------------------------------------------------------------------------------------------------------------

                #empiezacrud
            elif action == 'gruposcarrera': #ver grupoCarrera
                try:
                    url_vars= ''
                    data['title'] = 'Grupos de Carrera'
                    grupos = CarreraGrupo.objects.filter(status=True)
                    url_vars += "&action={}".format(action)
                    #dentro del if es buscar
                    if 'search' in request.GET:
                        data['search'] = search = request.GET['search']
                        url_vars += "&search={}".format(search)
                        grupos = grupos.filter(nombre__icontains=search)

                    paging = MiPaginador(grupos, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['listado'] = page.object_list
                    data['list_count'] = len(grupos)
                    data['url_vars'] = url_vars
                    return render(request, "empresa/gruposcarrera.html", data)
                except Exception as ex:
                    pass
            elif action == 'solicitudes':
                try:
                    url_vars= ''
                    data['title'] = 'Solicitudes de validación titulo'
                    grupos = SolicitudRevisionTitulo.objects.filter(status=True)
                    url_vars += "&action={}".format(action)
                    #dentro del if es buscar
                    if 'search' in request.GET:
                        data['search'] = search = request.GET['search']
                        url_vars += "&search={}".format(search)
                        grupos = grupos.filter(Q(inscripcion__persona__nombres__icontains=search) | Q(inscripcion__persona__apellido1__icontains=search)
                                               | Q(inscripcion__persona__apellido2__icontains=search) | Q(inscripcion__persona__ceula__icontains=search))
                    grupos = grupos.order_by('-fecha_creacion')
                    paging = MiPaginador(grupos, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['listado'] = page.object_list
                    data['list_count'] = len(grupos)
                    data['url_vars'] = url_vars
                    return render(request, "empresa/solicitudes.html", data)
                except Exception as ex:
                    pass
            elif action == 'gestionarsolicitud':
                try:
                    from empleo.forms import GestionarSolicitudRevisionTituloForm
                    data['solicitud'] = solicitud = SolicitudRevisionTitulo.objects.filter(status=True, id=int(encrypt(request.GET['id']))).first()
                    if not solicitud:
                        raise NameError('Solicitud no encontrada')
                    data['action'] = action
                    form = GestionarSolicitudRevisionTituloForm()
                    data['form'] = form
                    template = get_template("empleo/postular/gestionarsolicitud.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            return HttpResponseRedirect(request.path)

        else:
            try:
                data = {}
                adduserdata(request, data)
                data['title'] = u'Listado de empresas inscritas'
                search = None
                ids = None
                estado = request.GET.get('estado', '')
                url_vars = ''
                filtro = (Q(status=True))
                conofertas = OfertaLaboralEmpresa.objects.filter(status=True).values_list('empresa_id', flat=True)
                solicitudes = SolicitudAprobacionEmpresa.objects.filter(status=True, tiposolicitud=0).values_list('empleador_id', flat=True)
                empresas = Empleador.objects.filter(Q(id__in=conofertas) | Q(id__in=solicitudes)).order_by('solicitudaprobacionempresa__estadoempresa','fecha_creacion')
                if estado:
                    data['estadoapto'] = int(estado)
                    url_vars += "&estado={}".format(estado)
                    empresas = empresas.filter(Q(solicitudaprobacionempresa__estadoempresa=estado))
                if 's' in request.GET:
                    data['s'] = s = request.GET['s']
                    url_vars += "&s={}".format(s)
                    filtro = filtro & (Q(persona__apellido1__icontains=s) | Q(persona__apellido2__icontains=s) | Q(persona__nombres__icontains=s) | Q(nombrecorto__icontains=s))
                data["url_vars"] = url_vars
                empresas = empresas.filter(filtro)
                paging = MiPaginador(empresas, 25)
                p = 1
                try:
                    paginasesion = 1
                    if 'paginador' in request.session:
                        paginasesion = int(request.session['paginador'])
                    if 'page' in request.GET:
                        p = int(request.GET['page'])
                    else:
                        p = paginasesion
                    try:
                        page = paging.page(p)
                    except:
                        p = 1
                    page = paging.page(p)
                except:
                    page = paging.page(p)

                request.session['paginador'] = p
                data['paging'] = paging
                data['rangospaging'] = paging.rangos_paginado(p)
                data['page'] = page
                data['search'] = search if search else ""
                data['total']=len(empresas)
                data['empresas'] = page.object_list
                data['sbu_a'] = variable_valor('SBU_VALOR')
                data['estado'] = ESTADOS_EMPRESA

                #Cantidad de registros para los cuadritos estadisticos
                nDays = 7  # To get records of
                lastNDays = hoy - timedelta(nDays)

                cEmpresas = Empleador.objects.filter(solicitudaprobacionempresa__estadoempresa=1, status=True)
                numempresas = {'count': cEmpresas.count(),
                               'last_records': cEmpresas.filter(fecha_creacion__gte=lastNDays).count()}

                cOfertas = OfertaLaboralEmpresa.objects.filter(empresa__solicitudaprobacionempresa__estadoempresa=1,
                                                               empresa__status=True, estadooferta=1, status=True)
                numofertas = {'count': cOfertas.count(),
                              'last_records': cOfertas.filter(fecha_creacion__gte=lastNDays).count()}

                numofertasdisp = {
                    'count': cOfertas.filter(finiciopostulacion__lte=hoy, ffinpostlacion__gte=hoy).count(),
                    'last_records': cOfertas.filter(finiciopostulacion__lte=hoy, ffinpostlacion__gte=hoy,
                                                    fecha_creacion__gte=lastNDays).count()}

                cPostulantes = PersonaAplicaOferta.objects.filter(oferta__status=True, oferta__empresa__status=True,
                                                                  status=True)
                numpostulantes = {'count': cPostulantes.count(),
                                  'last_records': cPostulantes.filter(fecha_creacion__gte=lastNDays).count()}

                ayer = hoy - timedelta(1)
                cUsuarios = LogEntryLogin.objects.filter(action_app=7, action_flag=1, action_time__date__lte=hoy)
                numusuarios = {'count': cUsuarios.count(),
                                  'last_records': cUsuarios.filter(action_time__date__exact=hoy).count()}

                data['numempresas'] = numempresas
                data['numofertas'] = numofertas
                data['numofertasdisp'] = numofertasdisp
                data['numpostulantes'] = numpostulantes
                data['numusuarios'] = numusuarios
                data['nDays'] = nDays

                if 'exportar_excel_detallado' in request.GET:
                    wb = openxl.Workbook()
                    wb["Sheet"].title = "Reporte_empresas_detallado"
                    ws = wb.active
                    style_title = openxlFont(name='Arial', size=16, bold=True)
                    style_cab = openxlFont(name='Arial', size=10, bold=True)
                    alinear = alin(horizontal="center", vertical="center")
                    response = HttpResponse(content_type="application/ms-excel")
                    response[
                        'Content-Disposition'] = 'attachment; filename=Reporte detallado de empresas inscritas' + '-' + random.randint(
                        1, 10000).__str__() + '.xlsx'
                    ws.column_dimensions['B'].width = 35
                    ws.column_dimensions['C'].width = 20
                    ws.column_dimensions['D'].width = 15
                    ws.column_dimensions['E'].width = 15
                    ws.column_dimensions['F'].width = 35
                    ws.column_dimensions['G'].width = 35
                    ws.column_dimensions['H'].width = 25
                    ws.column_dimensions['I'].width = 25
                    ws.column_dimensions['J'].width = 20
                    ws.column_dimensions['K'].width = 25
                    ws.merge_cells('A1:J1')
                    ws['A1'] = 'INFORMACION DETALLADA DE EMPRESAS INSCRITAS'
                    celda1 = ws['A1']
                    celda1.font = style_title
                    celda1.alignment = alinear

                    columns = [u"N°", u"NOMBRE EMPRESA", u"RUC", u"FEC. REGISTRO", u"ESTADO",
                               u"NOMBRE OFERTA", u"POSTULANTE/S",u"CEDULA POSTULANTE", u"ESTADO HOJA DE VIDA" , u"CONTRATADO"
                               ]
                    row_num = 3
                    for col_num in range(0, len(columns)):
                        celda = ws.cell(row=row_num, column=(col_num + 1), value=columns[col_num])
                        celda.font = style_cab
                    row_num = 4
                    numero=1
                    for list in empresas:
                        if list.total_ofertas_empleo() != 0:
                            ofertas = OfertaLaboralEmpresa.objects.filter(empresa_id=list.id, status=True)
                            for oferta in ofertas:
                                if oferta.total_postulantes() >0:
                                    postulantes = PersonaAplicaOferta.objects.filter(oferta_id=oferta.id, status=True)
                                    for postulante in postulantes:
                                        ws.cell(row=row_num, column=1, value=numero)
                                        ws.cell(row=row_num, column=2, value=str(list.nombre))
                                        ws.cell(row=row_num, column=3, value=str(list.persona.ruc))
                                        ws.cell(row=row_num, column=4, value=str(list.fecha_creacion.date()))
                                        ws.cell(row=row_num, column=5, value=str(
                                            list.estado_registro().get_estadoempresa_display()) if list.estado_registro() else "")
                                        ws.cell(row=row_num, column=6, value=str(oferta.titulo))
                                        ws.cell(row=row_num, column=7, value=str(postulante.persona.nombre_completo_minus()))
                                        ws.cell(row=row_num, column=8, value=str(postulante.persona.cedula))
                                        ws.cell(row=row_num, column=9, value=str(postulante.get_estado_display()))
                                        if postulante.estado == 2:
                                            if postulante.estcontrato == 0:
                                                estadocontrato = "Pendiente"
                                            if postulante.estcontrato == 1:
                                                estadocontrato = "Si"
                                            if postulante.estcontrato == 2:
                                                estadocontrato = "No"
                                        else:
                                            estadocontrato = "-"
                                        ws.cell(row=row_num, column=10, value=str(estadocontrato))
                                        row_num += 1
                                        numero += 1
                    wb.save(response)
                    return response

                if 'exportar_excel' in request.GET:
                    wb = openxl.Workbook()
                    wb["Sheet"].title = "Reporte_empresas"
                    ws = wb.active
                    style_title = openxlFont(name='Arial', size=16, bold=True)
                    style_cab = openxlFont(name='Arial', size=10, bold=True)
                    alinear = alin(horizontal="center", vertical="center")
                    response = HttpResponse(content_type="application/ms-excel")
                    response[
                        'Content-Disposition'] = 'attachment; filename=Reporte de empresas inscritas' + '-' + random.randint(
                        1, 10000).__str__() + '.xlsx'
                    ws.column_dimensions['B'].width = 25
                    ws.column_dimensions['C'].width = 25
                    ws.column_dimensions['D'].width = 15
                    ws.column_dimensions['E'].width = 25
                    ws.column_dimensions['F'].width = 25
                    ws.column_dimensions['G'].width = 25
                    ws.column_dimensions['H'].width = 25
                    ws.merge_cells('A1:J1')
                    ws['A1'] = 'EMPRESAS INSCRITAS'
                    celda1 = ws['A1']
                    celda1.font = style_title
                    celda1.alignment = alinear

                    columns = [u"N°", u"NOMBRE", u"RUC", u"TIPO", u"SECTOR ECONÓMICO",
                               u"UBICACIÓN", u"TELEFONO",u"EMAIL", u"ESTADO", u"TOTAL OFERTAS"
                               ]
                    row_num = 3
                    for col_num in range(0, len(columns)):
                        celda = ws.cell(row=row_num, column=(col_num + 1), value=columns[col_num])
                        celda.font = style_cab
                    row_num = 4
                    numero=1
                    for list in empresas:
                        ws.cell(row=row_num, column=1, value=numero)
                        ws.cell(row=row_num, column=2, value=str(list.nombre))
                        ws.cell(row=row_num, column=3, value=str(list.persona.ruc))
                        ws.cell(row=row_num, column=4, value=str(list.tipoempresa).capitalize())
                        ws.cell(row=row_num, column=5, value=str(list.get_sectoreconomico_display().capitalize()))
                        ws.cell(row=row_num, column=6, value=str(list.persona.direccion).capitalize())
                        ws.cell(row=row_num, column=7, value=str(list.persona.telefono))
                        ws.cell(row=row_num, column=8, value=str(list.persona.email))
                        ws.cell(row=row_num, column=9, value=str(list.estado_registro().get_estadoempresa_display()) if list.estado_registro() else "")
                        ws.cell(row=row_num, column=10, value=str(list.total_ofertas_empleo()))
                        row_num += 1
                        numero += 1
                    wb.save(response)
                    return response
                return render(request, "adm_ofertalaboral/view_empresas_unemi_empleo.html", data)
            except Exception as ex:
                print(ex)
            # SE COMENTO PARA QUE SE MUESTRE LA LISTA DE EMPRESAS PARA UNEMI EMPLEO
            # if 's' in request.GET:
            #     search = request.GET['s']
            #     ofertas = OfertaLaboral.objects.filter(Q(empresa__nombre__icontains=search) | Q(cargo__icontains=search) | Q(descripcion__icontains=search), status=True).order_by('-id')
            # elif 'id' in request.GET:
            #     ids = request.GET['id']
            #     ofertas = OfertaLaboral.objects.filter(id=ids,status=True).order_by('-id')
            # else:
            #     ofertas = OfertaLaboral.objects.filter(status=True).order_by('-id')
            # paging = MiPaginador(ofertas, 25)
            # p = 1
            # try:
            #     paginasesion = 1
            #     if 'paginador' in request.session:
            #         paginasesion = int(request.session['paginador'])
            #     if 'page' in request.GET:
            #         p = int(request.GET['page'])
            #     else:
            #         p = paginasesion
            #     try:
            #         page = paging.page(p)
            #     except:
            #         p = 1
            #     page = paging.page(p)
            # except:
            #     page = paging.page(p)
            #
            # request.session['paginador'] = p
            # data['paging'] = paging
            # data['rangospaging'] = paging.rangos_paginado(p)
            # data['page'] = page
            # data['search'] = search if search else ""
            # data['ids'] = ids if ids else ""
            # data['ofertas'] = page.object_list
            # data['reporte_1'] = obtener_reporte('oferta_laboral')
            # return render(request, "adm_ofertalaboral/view.html", data)
