# coding=utf-8
# !/usr/bin/env python

import os
import sys

# SITE_ROOT = os.path.dirname(os.path.realpath(__file__))

YOUR_PATH = os.path.dirname(os.path.realpath(__file__))
# print(f"YOUR_PATH: {YOUR_PATH}")
SITE_ROOT = os.path.dirname(os.path.dirname(YOUR_PATH))
SITE_ROOT = os.path.join(SITE_ROOT, '')
# print(f"SITE_ROOT: {SITE_ROOT}")
your_djangoproject_home = os.path.split(SITE_ROOT)[0]
# print(f"your_djangoproject_home: {your_djangoproject_home}")
sys.path.append(your_djangoproject_home)

import xlwt
from webpush import send_user_notification
from xlwt import easyxf
from django.core.wsgi import get_wsgi_application

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "settings")
application = get_wsgi_application()
from datetime import datetime, timedelta
from django.db import transaction
from sga.models import Materia, Clase, Leccion, LeccionGrupo, AsistenciaLeccion, CamposTitulosPostulacion
from sga.funciones import variable_valor, notificacion
# from sga.models import *
# from sagest.models import *
import xlrd
from postulate.models import Partida, PersonaAplicarPartida, PersonaFormacionAcademicoPartida, PartidaTribunal
from sga.models import *
from sagest.models import *
from django.db.models import Sum, F, FloatField, IntegerField
from django.db.models.functions import Coalesce
from settings import MEDIA_ROOT, BASE_DIR
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, colors
from django.http import HttpResponse
from sga.models import Modulo
from gdocumental.models import *
from bd.models import *
from moodle import moodle
from postulate.models import *
from core.firmar_documentos import verificarFirmasPDF


def empadronarpersonas2023milagro():
    cadena = ''
    linea, excluidos, conexito = 0, 0, 0
    try:
        archivo_ = 'MILAGRO_2023'
        # url_archivo = "{}/{}.xlsx".format(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), archivo_)
        url_archivo = "{}/media/{}.xlsx".format(SITE_STORAGE, archivo_)
        print('Leyendo....')
        # wb = openpyxl.load_workbook(filename="{}/media/{}.xlsx".format(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), archivo_))
        wb = openpyxl.load_workbook(filename=url_archivo)
        ws = wb.get_sheet_by_name("Hoja2")
        # worksheet = wb["Listado"]
        worksheet = ws
        lis_excluidos = []
        print('Iniciando....')
        linea_archivo = 1
        ids_excluir_a_rechazar = []
        periodo = CabPadronElectoral.objects.get(pk=3, status=True)
        for row in worksheet.iter_rows(min_row=0):
            if linea >= 1:
                currentValues, cadena = [], ''
                for cell in row:
                    cadena += str(cell.value) + ' '
                    currentValues.append(str(cell.value))
                cedula_ = currentValues[1]
                coordinacion = currentValues[3]
                mesa = currentValues[4]
                nombre_mesa = f"{mesa} {coordinacion} - MILAGRO"
                qsmesa = MesasPadronElectoral.objects.filter(status=True, periodo=periodo, orden=1, nombre=nombre_mesa)
                mesa_ = None
                if qsmesa.exists():
                    mesa_ = qsmesa.first()
                else:
                    mesa_ = MesasPadronElectoral(periodo=periodo, orden=1, nombre=nombre_mesa)
                    mesa_.save()

                if mesa_:
                    qsempadronada = DetPersonaPadronElectoral.objects.filter(status=True, cab=periodo, persona__cedula__icontains=cedula_, tipo=1)
                    empadronado = None
                    if qsempadronada.exists():
                        empadronado = qsempadronada.first()
                        empadronado.lugar = nombre_mesa
                        empadronado.mesa = mesa_
                        empadronado.save()
                    else:
                        lis_excluidos.append({'cedula': cedula_, 'obs': f'No existe empadronado'})
                        excluidos += 1
                print("Linea {}/? - Persona: {}".format(linea, cedula_))
            linea += 1
            linea_archivo += 1
        print('Total Leidos con exito: {}'.format(conexito))
        print('Total Excluidos: {}'.format(excluidos))
        print(lis_excluidos)
    except Exception as ex:
        textoerror = '{} Linea:{} - Info: {} / Lectura: {}'.format(str(ex), sys.exc_info()[-1].tb_lineno, cadena, linea)
        print(textoerror)


empadronarpersonas2023milagro()