import os
import sys
import io
import xlsxwriter
import xlwt
import openpyxl
import xlwt
from xlwt import *
from django.http import HttpResponse
from xlwt import *

# SITE_ROOT = os.path.dirname(os.path.realpath(__file__))
YOUR_PATH = os.path.dirname(os.path.realpath(__file__))
# print(f"YOUR_PATH: {YOUR_PATH}")
SITE_ROOT = os.path.dirname(os.path.dirname(YOUR_PATH))
SITE_ROOT = os.path.join(SITE_ROOT, '')
# print(f"SITE_ROOT: {SITE_ROOT}")
your_djangoproject_home = os.path.split(SITE_ROOT)[0]
# print(f"your_djangoproject_home: {your_djangoproject_home}")
sys.path.append(your_djangoproject_home)

from django.core.wsgi import get_wsgi_application

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "settings")
application = get_wsgi_application()

from django.http import HttpResponse
from settings import MEDIA_ROOT, BASE_DIR
from xlwt import easyxf, XFStyle
from sga.adm_criteriosactividadesdocente import asistencia_tutoria
from inno.models import *
from sga.models import *
from sagest.models import *
from balcon.models import *
from inno.funciones import *
from sga.funciones import  notificacion
from Moodle_Funciones import crearhtmlphpmoodle
import concurrent.futures

def cerrar_materias_transversales_3():
    materias = Materia.objects.filter(status=True, nivel__periodo_id=177, cerrado=False,
                                      asignaturamalla__malla__carrera__coordinacion__in=[3], modeloevaluativo_id=27)
    for materia in materias:
        print(materia)
        for asig in materia.asignados_a_esta_materia():
            asig.cerrado = True
            asig.save(actualiza=False)
            asig.actualiza_estado()
        for asig in materia.asignados_a_esta_materia():
            asig.cierre_materia_asignada()

        materia.cerrado = True
        materia.fechacierre = datetime.now().date()
        materia.save()

#cerrar_materias_transversales_3()

def actualizar_nivel_inscripcion_malla2():
    matriculas = Matricula.objects.filter(status=True, nivel__periodo_id=177, inscripcion__carrera_id__in=[131,129] )
    for matricula in matriculas:
        inscripcion = matricula.inscripcion
        print('ACTUALIZANDO- ', inscripcion.persona.cedula)
        inscripcion.actualizar_nivel()
        print('ACTUALIZADO')
    print('FIN')

#actualizar_nivel_inscripcion_malla2()

def calificacion_transversales_en_linea():
    try:
        periodo = Periodo.objects.get(id=177)
        asignaturas = DetalleGrupoAsignatura.objects.values_list('asignatura_id', flat=True).filter(status=True,
                                                                                                    grupo_id__in=[1, 2,
                                                                                                                  3])

        materias = Materia.objects.filter(status=True, nivel__periodo_id=177,
                                          asignaturamalla__asignatura_id__in=asignaturas,
                                          modeloevaluativo_id=27, asignaturamalla__malla__carrera__coordinacion__in=[3])

        for materia in materias:
            idcursomoodle = materia.idcursomoodle
            materiasasignadas = MateriaAsignada.objects.filter(status=True,
                                                               matricula__nivel__periodo=periodo,
                                                               materia=materia,
                                                               matricula__bloqueomatricula=False,
                                                               matricula__retiradomatricula=False, materia__status=True,
                                                               matricula__status=True,
                                                               matricula__inscripcion__carrera__modalidad__in=[1, 2, 3])

            cursor = connections['moodle_db'].cursor()
            for materiaasignada in materiasasignadas:
                # guardo_nota = False
                usuario = materiaasignada.matricula.inscripcion.persona.usuario.username
                # SE NECESITA EL ID DE CURSO MOODLE
                sql = """
                                                    SELECT ROUND(nota.finalgrade,2), UPPER(gc.fullname)
                                                            FROM mooc_grade_grades nota
                                                    INNER JOIN mooc_grade_items it ON nota.itemid=it.id AND courseid=%s AND itemtype='category'
                                                    INNER JOIN mooc_grade_categories gc ON gc.courseid=it.courseid AND gc.id=it.iteminstance AND gc.depth=2
                                                    INNER JOIN mooc_user us ON nota.userid=us.id
                                                    WHERE us.username = '%s' and UPPER(gc.fullname)='RE'
                                                    ORDER BY it.sortorder
                                                """ % (str(idcursomoodle), usuario)

                cursor.execute(sql)
                results = cursor.fetchall()
                if results:
                    for notasmooc in results:
                        campo = materiaasignada.campo(notasmooc[1].upper())
                        if not campo:
                            print('revisar curso moodle - ', materiaasignada.materia.id, 'idcursomoodle -',
                                  materiaasignada.materia.idcursomoodle)
                            continue
                        if type(notasmooc[0]) is Decimal:
                            if null_to_decimal(campo.valor) != float(notasmooc[0]):
                                actualizar_nota_planificacion(materiaasignada.id, notasmooc[1].upper(), notasmooc[0])
                                auditorianotas = AuditoriaNotas(evaluaciongenerica=campo, manual=False,
                                                                calificacion=notasmooc[0])
                                auditorianotas.save()
                                print('importacion exitosa - ', materiaasignada)

                        else:
                            if null_to_decimal(campo.valor) != float(0):
                                actualizar_nota_planificacion(materiaasignada.id, notasmooc[1].upper(), notasmooc[0])
                                auditorianotas = AuditoriaNotas(evaluaciongenerica=campo, manual=False, calificacion=0)
                                auditorianotas.save()
                                print('importacion exitosa - ', materiaasignada)

                else:
                    detallemodelo = DetalleModeloEvaluativo.objects.get(pk=125)
                    campo = materiaasignada.campo(detallemodelo.nombre)
                    actualizar_nota_planificacion(materiaasignada.id, detallemodelo.nombre, 0)
                    auditorianotas = AuditoriaNotas(evaluaciongenerica=campo, manual=False, calificacion=0)
                    auditorianotas.save()

            for asig in materia.asignados_a_esta_materia():
                asig.cerrado = True
                asig.save(actualiza=False)
                asig.actualiza_estado()
                asig.cierre_materia_asignada()

            print('PROCESO FINALIZADO')






    except Exception as ex:
        msg = ex.__str__()

        textoerror = '{} Linea:{}'.format(str(ex), sys.exc_info()[-1].tb_lineno)
        print(textoerror)
        print(msg)


#calificacion_transversales_en_linea()

@transaction.atomic()
def homologacion():

    #verificar que todos los estudiantes tengan la misma malla
    cadena = ''
    linea, excluidos, conexito = 0, 0, 0
    try:
        libre_origen = '/homologacion_derecho.xls'
        fuentecabecera = easyxf(
            'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
        fuentenormal = easyxf(
            'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
        output_folder = MEDIA_ROOT
        output_folder = os.path.join(os.path.join(BASE_DIR))
        # liborigen = xlrd.open_workbook(output_folder + libre_origen)
        libdestino = xlwt.Workbook()
        hojadestino = libdestino.add_sheet('HOJA1')
        fil = 0
        columnas = [(u"CEDULA", 6000),
                   (u"APELLIDOS Y NOMBRES", 6000),
                   (u"OBSERVACIÓN", 6000)
        ]
        for col_num in range(len(columnas)):
            hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
            hojadestino.col(col_num).width = columnas[col_num][1]
        fila = 1
        titulacion = 0

        lin = 0
        miarchivo = openpyxl.load_workbook("CEDULAS_DERECHO_final.xlsx")
        #miarchivo = openpyxl.load_workbook("prueba_2.xlsx")

        ws = miarchivo.get_sheet_by_name("Hoja1")
        worksheet = ws
        c = 0
        cont = 0
        periodo_id=177
        carrera_id=126
        mallaantigua_id=207
        mallanueva_id=480

        for row in worksheet.iter_rows(min_row=0):
            if lin >= 0:
                currentValues, cadena = [], ''
                for cell in row:
                    cadena += str(cell.value) + ' '
                    currentValues.append(str(cell.value))
                identificacion = currentValues[0]

                matricula = Matricula.objects.filter(status=True, inscripcion__carrera__id=carrera_id, inscripcion__persona__cedula=identificacion).first()
                cont += 1
                matricula.pasoayuda = True
                matricula.save()
                print(u"%s - %s" % (matricula, cont))
                inscripcion = matricula.inscripcion
                hojadestino.write(fila, 0, matricula.inscripcion.persona.identificacion(), fuentenormal)
                hojadestino.write(fila, 1, matricula.inscripcion.persona.nombre_completo(), fuentenormal)
                hojadestino.write(fila, 2, matricula.nivel.periodo.nombre, fuentenormal)

                if InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallaantigua_id).exists():
                    imantigua = InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallaantigua_id)[0]
                    imantigua.status = False
                    imantigua.save()
                    print(u"Desactiva antigua inscripcion -----------------------------")

                if not InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallanueva_id).exists():
                    imnueva = InscripcionMalla(inscripcion=inscripcion, malla_id=mallanueva_id)
                    imnueva.save()
                    print(u"Crea nueva inscripcion -----------------------------")

                equivalencias = TablaEquivalenciaAsignaturas.objects.filter(status=True, asignaturamalla__malla_id=mallaantigua_id).order_by('asignaturamallasalto__nivelmalla__orden')
                for equivalencia in equivalencias:
                    print(u"nueva - %s" % equivalencia.asignaturamallasalto)
                    recordantiguo = inscripcion.recordacademico_set.filter(status=True,asignaturamalla_id=equivalencia.asignaturamalla_id).first()

                    if recordantiguo:
                        print(u"anterior - %s" % equivalencia.asignaturamalla)
                        print(u"Record antiguo: %s" % recordantiguo)
                        recordnuevo = None
                        recordantiguo.status = False
                        recordantiguo.save(update_asignaturamalla=False)

                        if equivalencia.asignaturamallasalto_id == 10623 or equivalencia.asignaturamallasalto_id == 10627 or equivalencia.asignaturamallasalto_id == 10615:
                            observaciones = recordantiguo.observaciones + " Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15"
                            homologada = True
                        else:
                            observaciones = recordantiguo.observaciones + " Migración con base a RESOLUCIÓN CGA-SO-9-2023-NO15 Y RESOLUCIÓN OCS-SO-18-2023-NO14"
                            homologada = False
                        if not RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                              asignaturamalla=equivalencia.asignaturamallasalto).exists():



                            recordnuevo = RecordAcademico(inscripcion=inscripcion,
                                                          matriculas=recordantiguo.matriculas,
                                                          asignaturamalla=equivalencia.asignaturamallasalto,
                                                          asignatura=equivalencia.asignaturamallasalto.asignatura,
                                                          asignaturaold_id=recordantiguo.asignatura.id,
                                                          nota=recordantiguo.nota,
                                                          asistencia=recordantiguo.asistencia,
                                                          sinasistencia=recordantiguo.sinasistencia,
                                                          fecha=recordantiguo.fecha,
                                                          noaplica=recordantiguo.noaplica,
                                                          aprobada=recordantiguo.aprobada,
                                                          convalidacion=recordantiguo.convalidacion,
                                                          pendiente=recordantiguo.pendiente,
                                                          creditos=equivalencia.asignaturamallasalto.creditos,
                                                          horas=equivalencia.asignaturamallasalto.horas,
                                                          valida=recordantiguo.valida,
                                                          validapromedio=recordantiguo.validapromedio,
                                                          observaciones=observaciones,
                                                          homologada=homologada,
                                                          materiaregular=recordantiguo.materiaregular,
                                                          materiacurso=None,
                                                          completonota=recordantiguo.completonota,
                                                          completoasistencia=recordantiguo.completoasistencia,
                                                          fechainicio=recordantiguo.fechainicio,
                                                          fechafin=recordantiguo.fechafin,
                                                          suficiencia=recordantiguo.suficiencia,
                                                          asignaturamallahistorico_id=recordantiguo.asignaturamalla.id,
                                                          reverso=False)
                            recordnuevo.save()
                            print(u"Crea nuevo record %s" % recordnuevo)


                        elif RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                            asignaturamalla=equivalencia.asignaturamallasalto):
                            recordnuevo = RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                         asignaturamalla=equivalencia.asignaturamallasalto)[0]
                            recordnuevo.matriculas = recordantiguo.matriculas
                            recordnuevo.asignaturamalla = equivalencia.asignaturamallasalto
                            recordnuevo.asignatura = equivalencia.asignaturamallasalto.asignatura
                            recordnuevo.asignaturaold = recordantiguo.asignaturamalla.asignatura
                            recordnuevo.nota = recordantiguo.nota
                            recordnuevo.asistencia = recordantiguo.asistencia
                            recordnuevo.sinasistencia = recordantiguo.sinasistencia
                            recordnuevo.fecha = recordantiguo.fecha
                            recordnuevo.noaplica = recordantiguo.noaplica
                            recordnuevo.aprobada = recordantiguo.aprobada
                            recordnuevo.convalidacion = recordantiguo.convalidacion
                            recordnuevo.pendiente = recordantiguo.pendiente
                            recordnuevo.creditos = equivalencia.asignaturamallasalto.creditos
                            recordnuevo.horas = equivalencia.asignaturamallasalto.horas
                            recordnuevo.valida = recordantiguo.valida
                            recordnuevo.validapromedio = recordantiguo.validapromedio
                            recordnuevo.observaciones = observaciones
                            recordnuevo.homologada = homologada
                            recordnuevo.materiaregular = recordantiguo.materiaregular
                            recordnuevo.materiacurso = None
                            recordnuevo.completonota = recordantiguo.completonota
                            recordnuevo.completoasistencia = recordantiguo.completoasistencia
                            recordnuevo.fechainicio = recordantiguo.fechainicio
                            recordnuevo.fechafin = recordantiguo.fechafin
                            recordnuevo.suficiencia = recordantiguo.suficiencia
                            recordnuevo.asignaturamallahistorico = recordantiguo.asignaturamalla
                            recordnuevo.reverso = False
                            recordnuevo.save()

                        if recordnuevo:
                            historicos = HistoricoRecordAcademico.objects.filter(status=True,
                                                                                 recordacademico=recordantiguo).update(recordacademico=recordnuevo,
                                                                                                                       creditos=recordnuevo.creditos,
                                                                                                                       horas=recordnuevo.horas,
                                                                                                                       homologada=recordnuevo.homologada)
                            respaldo = RespaldoRecordAcademico.objects.filter(status=True,recordacademicooriginal=recordantiguo)

                            if not respaldo.exists():
                                respaldorecord = RespaldoRecordAcademico(
                                    recordacademicooriginal=recordantiguo,
                                    recordacademiconuevo=recordnuevo
                                )
                                respaldorecord.save()
                            else:
                                respaldorecord = respaldo[0]
                                respaldorecord.recordacademiconuevo = recordnuevo
                                respaldorecord.save()
                            print(u"Record actualizado %s" % recordnuevo)


                    else:
                        hojadestino.write(fila, 3, "NO ENCONTRO RECORD ANTIGUO %s" % equivalencia.asignaturamalla)
                        fila += 1

                fila += 1

                time.sleep(3)

            lin += 1

        libdestino.save(output_folder + libre_origen)
        print(output_folder + libre_origen)
        print("Proceso finalizado. . .")


    except Exception as ex:
        transaction.set_rollback(True)
        print('error: %s' % ex)
        hojadestino.write(fila, 3, str(ex))
        fila += 1


def llenar_tabla_equivalencias():
    try:
        miarchivo = openpyxl.load_workbook("equivalencia_malla_derecho_prueba.xlsx")
        lista = miarchivo.get_sheet_by_name('MALLA_NUEVA')
        totallista = lista.rows
        a=0
        for filas in totallista:
            a += 1
            if a > 1:
                idasignaturamallaanterior = int(filas[4].value)
                idasignaturamallanueva=int(filas[1].value)
                if not TablaEquivalenciaAsignaturas.objects.filter(status=True,asignaturamalla_id=idasignaturamallaanterior).exists():
                    tablaeq=TablaEquivalenciaAsignaturas(asignaturamalla_id=idasignaturamallaanterior,
                                                 asignaturamallasalto_id=idasignaturamallanueva)
                    tablaeq.save()
                    print(u"INSERTA EQUIVALENCIA %s"%tablaeq)
                else:
                    tablaeq=TablaEquivalenciaAsignaturas.objects.filter(status=True, asignaturamalla_id=idasignaturamallaanterior)[0]
                    tablaeq.asignaturamallasalto_id=idasignaturamallanueva
                    tablaeq.save()
                    print(u"ACTUALIZA EQUIVALENCIA %s" % tablaeq)
                print(u"Fila %s"%a)
        # homologacion()
    except Exception as ex:
            print('error: %s' % ex)

@transaction.atomic()
def homologacion_psicologia():
    # Verificar que todos los estudiantes tengan la misma malla
    cadena = ''
    linea, excluidos, conexito = 0, 0, 0
    try:
        libre_origen = '/homologacion_psicologia.xls'
        fuentecabecera = easyxf('font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
        fuentenormal = easyxf('font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
        output_folder = MEDIA_ROOT
        output_folder = os.path.join(os.path.join(BASE_DIR))
        # liborigen = xlrd.open_workbook(output_folder + libre_origen)
        libdestino = xlwt.Workbook()
        hojadestino = libdestino.add_sheet('HOJA1')
        fil = 0
        columnas = [(u"CEDULA", 6000), (u"APELLIDOS Y NOMBRES", 6000), (u"OBSERVACIÓN", 6000)]
        for col_num in range(len(columnas)):
            hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
            hojadestino.col(col_num).width = columnas[col_num][1]
        fila = 1
        titulacion = 0

        lin = 0
        miarchivo = openpyxl.load_workbook("CEDULAS_PSICOLOGIA_prueba.xlsx")
        ws = miarchivo.get_sheet_by_name("Hoja1")
        worksheet = ws
        c = 0
        cont = 0
        periodo_id=177
        carrera_id=132
        mallaantigua_id=204
        mallanueva_id=479
        for row in worksheet.iter_rows(min_row=0):
            if lin >= 0:
                currentValues, cadena = [], ''
                for cell in row:
                    cadena += str(cell.value) + ' '
                    currentValues.append(str(cell.value))
                identificacion = currentValues[0]
                if not identificacion.isdigit():
                    return 0
                matricula = Matricula.objects.filter(status=True, inscripcion__carrera__id=carrera_id, inscripcion__persona__cedula=identificacion).first()
                cont += 1
                matricula.pasoayuda = True
                matricula.save()
                print(u"%s - %s" % (matricula, cont))
                inscripcion = matricula.inscripcion
                hojadestino.write(fila, 0, matricula.inscripcion.persona.identificacion(), fuentenormal)
                hojadestino.write(fila, 1, matricula.inscripcion.persona.nombre_completo(), fuentenormal)
                hojadestino.write(fila, 2, matricula.nivel.periodo.nombre, fuentenormal)

                if InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallaantigua_id).exists():
                    imantigua = InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallaantigua_id)[0]
                    imantigua.status = False
                    imantigua.save()
                    print(u"Desactiva antigua inscripcion -----------------------------")

                if not InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallanueva_id).exists():
                    imnueva = InscripcionMalla(inscripcion=inscripcion, malla_id=mallanueva_id)
                    imnueva.save()
                    print(u"Crea nueva inscripcion -----------------------------")

                equivalencias = TablaEquivalenciaAsignaturas.objects.filter(status=True, asignaturamalla__malla_id=mallaantigua_id).order_by('asignaturamallasalto__nivelmalla__orden')

                # Se crea una lista temporal para verificar si existen registros antes de modificarlos
                temporal = []
                for e in equivalencias:
                    temporal.append([inscripcion.recordacademico_set.filter(status=True, asignaturamalla=e.asignaturamalla).first(), e])

                for t in temporal:
                    recordantiguo, equivalencia = t

                    old, new = equivalencia.asignaturamalla, equivalencia.asignaturamallasalto
                    print(f"Nueva - {equivalencia.asignaturamallasalto}")

                    if recordantiguo:
                        print(f"Anterior - {equivalencia.asignaturamalla}")
                        print(f"Record antiguo: {recordantiguo}")

                        recordantiguo.status = False
                        recordantiguo.save(update_asignaturamalla=False)

                        recordnuevo, homologada = None, False
                        if equivalencia.asignaturamallasalto_id in (10646, 10649, 10654, 10639):
                            observaciones = recordantiguo.observaciones + " Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15"
                            homologada = True
                        else:
                            observaciones = recordantiguo.observaciones + " Migración con base a RESOLUCIÓN CGA-SO-9-2023-NO15 Y RESOLUCIÓN OCS-SO-18-2023-NO14"

                        if recordnuevo := inscripcion.recordacademico_set.filter(asignaturamalla=new, status=True).first():
                            recordnuevo.matriculas = recordantiguo.matriculas
                            recordnuevo.asignaturamalla = equivalencia.asignaturamallasalto
                            recordnuevo.asignatura = equivalencia.asignaturamallasalto.asignatura
                            recordnuevo.asignaturaold = recordantiguo.asignaturamalla.asignatura
                            recordnuevo.nota = recordantiguo.nota
                            recordnuevo.asistencia = recordantiguo.asistencia
                            recordnuevo.sinasistencia = recordantiguo.sinasistencia
                            recordnuevo.fecha = recordantiguo.fecha
                            recordnuevo.noaplica = recordantiguo.noaplica
                            recordnuevo.aprobada = recordantiguo.aprobada
                            recordnuevo.convalidacion = recordantiguo.convalidacion
                            recordnuevo.pendiente = recordantiguo.pendiente
                            recordnuevo.creditos = equivalencia.asignaturamallasalto.creditos
                            recordnuevo.horas = equivalencia.asignaturamallasalto.horas
                            recordnuevo.valida = recordantiguo.valida
                            recordnuevo.validapromedio = recordantiguo.validapromedio
                            recordnuevo.observaciones = observaciones
                            recordnuevo.homologada = homologada
                            recordnuevo.materiaregular = recordantiguo.materiaregular
                            recordnuevo.materiacurso = None
                            recordnuevo.completonota = recordantiguo.completonota
                            recordnuevo.completoasistencia = recordantiguo.completoasistencia
                            recordnuevo.fechainicio = recordantiguo.fechainicio
                            recordnuevo.fechafin = recordantiguo.fechafin
                            recordnuevo.suficiencia = recordantiguo.suficiencia
                            recordnuevo.asignaturamallahistorico = recordantiguo.asignaturamalla
                            recordnuevo.reverso = False
                            recordnuevo.save()
                        else:
                            recordnuevo = RecordAcademico(inscripcion=inscripcion,
                                                          matriculas=recordantiguo.matriculas,
                                                          asignaturamalla=equivalencia.asignaturamallasalto,
                                                          asignatura=equivalencia.asignaturamallasalto.asignatura,
                                                          asignaturaold_id=recordantiguo.asignatura.id,
                                                          nota=recordantiguo.nota,
                                                          asistencia=recordantiguo.asistencia,
                                                          sinasistencia=recordantiguo.sinasistencia,
                                                          fecha=recordantiguo.fecha,
                                                          noaplica=recordantiguo.noaplica,
                                                          aprobada=recordantiguo.aprobada,
                                                          convalidacion=recordantiguo.convalidacion,
                                                          pendiente=recordantiguo.pendiente,
                                                          creditos=equivalencia.asignaturamallasalto.creditos,
                                                          horas=equivalencia.asignaturamallasalto.horas,
                                                          valida=recordantiguo.valida,
                                                          validapromedio=recordantiguo.validapromedio,
                                                          observaciones=observaciones,
                                                          homologada=homologada,
                                                          materiaregular=recordantiguo.materiaregular,
                                                          materiacurso=None,
                                                          completonota=recordantiguo.completonota,
                                                          completoasistencia=recordantiguo.completoasistencia,
                                                          fechainicio=recordantiguo.fechainicio,
                                                          fechafin=recordantiguo.fechafin,
                                                          suficiencia=recordantiguo.suficiencia,
                                                          asignaturamallahistorico_id=recordantiguo.asignaturamalla.id,
                                                          reverso=False)
                            recordnuevo.save()
                            print(u"Crea nuevo record %s" % recordnuevo)

                        if recordnuevo:
                            historicos = HistoricoRecordAcademico.objects.filter(status=True, recordacademico=recordantiguo).update(recordacademico=recordnuevo, creditos=recordnuevo.creditos, horas=recordnuevo.horas, homologada=recordnuevo.homologada)
                            respaldo = RespaldoRecordAcademico.objects.filter(status=True,recordacademicooriginal=recordantiguo)

                            if not respaldo.exists():
                                respaldorecord = RespaldoRecordAcademico(recordacademicooriginal=recordantiguo, recordacademiconuevo=recordnuevo)
                                respaldorecord.save()
                            else:
                                respaldorecord = respaldo[0]
                                respaldorecord.recordacademiconuevo = recordnuevo
                                respaldorecord.save()
                            print(u"Record actualizado %s" % recordnuevo)
                    else:
                        hojadestino.write(fila, 3, "NO ENCONTRO RECORD ANTIGUO %s" % equivalencia.asignaturamalla)
                        fila += 1

                fila += 1

                time.sleep(3)

            lin += 1

        libdestino.save(output_folder + libre_origen)
        print(output_folder + libre_origen)
        print("Proceso finalizado. . .")
    except Exception as ex:
        transaction.set_rollback(True)
        print('error: %s' % ex)
        hojadestino.write(fila, 3, str(ex))
        fila += 1

def llenar_tabla_equivalencias_psicologia():
    try:
        miarchivo = openpyxl.load_workbook("matriz_equivalencia_psicologia.xlsx")
        lista = miarchivo.get_sheet_by_name('MALLA_NUEVA')
        totallista = lista.rows
        a=0
        for filas in totallista:
            a += 1
            if a > 1 and f"{filas[4].value}".isdigit():
                idasignaturamallaanterior, idasignaturamallanueva = int(filas[4].value), int(filas[1].value)
                msg = ''
                if tablaeq := TablaEquivalenciaAsignaturas.objects.filter(asignaturamallasalto_id=idasignaturamallanueva, asignaturamalla_id=idasignaturamallaanterior).first():
                    tablaeq.status = True
                    msg = u"ACTUALIZA EQUIVALENCIA %s"
                else:
                    tablaeq = TablaEquivalenciaAsignaturas(asignaturamalla_id=idasignaturamallaanterior, asignaturamallasalto_id=idasignaturamallanueva)
                    msg = u"INSERTA EQUIVALENCIA %s"

                tablaeq.save()
                print(msg % f'{tablaeq.asignaturamallasalto} - {tablaeq.asignaturamalla}')

                print(u"Fila %s"%a)
        # homologacion()
    except Exception as ex:
            print('error: %s' % ex)

@transaction.atomic()
def homologacion_economia():

    #verificar que todos los estudiantes tengan la misma malla
    cadena = ''
    linea, excluidos, conexito = 0, 0, 0
    try:
        libre_origen = '/homologacion_economia.xls'
        fuentecabecera = easyxf('font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
        fuentenormal = easyxf('font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
        output_folder = MEDIA_ROOT
        output_folder = os.path.join(os.path.join(BASE_DIR))
        # liborigen = xlrd.open_workbook(output_folder + libre_origen)
        libdestino = xlwt.Workbook()
        hojadestino = libdestino.add_sheet('HOJA1')
        fil = 0
        columnas = [(u"CEDULA", 6000),(u"APELLIDOS Y NOMBRES", 6000),(u"OBSERVACIÓN", 6000)]
        for col_num in range(len(columnas)):
            hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
            hojadestino.col(col_num).width = columnas[col_num][1]
        fila = 1
        titulacion = 0

        lin = 0
        miarchivo = openpyxl.load_workbook("cedulas_economia.xlsx")
        ws = miarchivo.get_sheet_by_name("Hoja1")
        worksheet = ws
        c = 0
        cont = 0
        periodo_id=177
        carrera_id=128
        mallaantigua_id=201
        mallanueva_id=489

        for row in worksheet.iter_rows(min_row=0):
            if lin >= 0:
                currentValues, cadena = [], ''
                for cell in row:
                    cadena += str(cell.value) + ' '
                    currentValues.append(str(cell.value))
                identificacion = currentValues[0]
                if not identificacion.isdigit():
                    return 0
                matricula = Matricula.objects.filter(status=True, inscripcion__carrera__id=carrera_id, inscripcion__persona__cedula=identificacion).first()
                cont += 1
                matricula.pasoayuda = True
                matricula.save()
                print(u"%s - %s" % (matricula, cont))
                inscripcion = matricula.inscripcion
                hojadestino.write(fila, 0, matricula.inscripcion.persona.identificacion(), fuentenormal)
                hojadestino.write(fila, 1, matricula.inscripcion.persona.nombre_completo(), fuentenormal)
                hojadestino.write(fila, 2, matricula.nivel.periodo.nombre, fuentenormal)

                if InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallaantigua_id).exists():
                    imantigua = InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallaantigua_id)[0]
                    imantigua.status = False
                    imantigua.save()
                    print(u"Desactiva antigua inscripcion -----------------------------")

                if not InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallanueva_id).exists():
                    imnueva = InscripcionMalla(inscripcion=inscripcion, malla_id=mallanueva_id)
                    imnueva.save()
                    print(u"Crea nueva inscripcion -----------------------------")

                equivalencias = TablaEquivalenciaAsignaturas.objects.filter(status=True, asignaturamalla__malla_id=mallaantigua_id).order_by('asignaturamallasalto__nivelmalla__orden')
                for equivalencia in equivalencias:
                    print(u"nueva - %s" % equivalencia.asignaturamallasalto)
                    recordantiguo = inscripcion.recordacademico_set.filter(status=True,asignaturamalla_id=equivalencia.asignaturamalla_id).first()

                    if recordantiguo:
                        print(u"anterior - %s" % equivalencia.asignaturamalla)
                        print(u"Record antiguo: %s" % recordantiguo)
                        recordnuevo = None
                        recordantiguo.status = False
                        recordantiguo.save(update_asignaturamalla=False)
                        if equivalencia.asignaturamallasalto_id in (10733, 10742, 10770, 10774):
                            observaciones = recordantiguo.observaciones + " Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15"
                            homologada = True
                        else:
                            observaciones = recordantiguo.observaciones + " Migración con base a RESOLUCIÓN CGA-SO-9-2023-NO15 Y RESOLUCIÓN OCS-SO-18-2023-NO14"
                            homologada = False
                        if not RecordAcademico.objects.filter(status=True, inscripcion=inscripcion, asignaturamalla=equivalencia.asignaturamallasalto).exists():
                            recordnuevo = RecordAcademico(inscripcion=inscripcion,
                                                          matriculas=recordantiguo.matriculas,
                                                          asignaturamalla=equivalencia.asignaturamallasalto,
                                                          asignatura=equivalencia.asignaturamallasalto.asignatura,
                                                          asignaturaold_id=recordantiguo.asignatura.id,
                                                          nota=recordantiguo.nota,
                                                          asistencia=recordantiguo.asistencia,
                                                          sinasistencia=recordantiguo.sinasistencia,
                                                          fecha=recordantiguo.fecha,
                                                          noaplica=recordantiguo.noaplica,
                                                          aprobada=recordantiguo.aprobada,
                                                          convalidacion=recordantiguo.convalidacion,
                                                          pendiente=recordantiguo.pendiente,
                                                          creditos=equivalencia.asignaturamallasalto.creditos,
                                                          horas=equivalencia.asignaturamallasalto.horas,
                                                          valida=recordantiguo.valida,
                                                          validapromedio=recordantiguo.validapromedio,
                                                          observaciones=observaciones,
                                                          homologada=homologada,
                                                          materiaregular=recordantiguo.materiaregular,
                                                          materiacurso=None,
                                                          completonota=recordantiguo.completonota,
                                                          completoasistencia=recordantiguo.completoasistencia,
                                                          fechainicio=recordantiguo.fechainicio,
                                                          fechafin=recordantiguo.fechafin,
                                                          suficiencia=recordantiguo.suficiencia,
                                                          asignaturamallahistorico_id=recordantiguo.asignaturamalla.id,
                                                          reverso=False)
                            recordnuevo.save()
                            print(u"Crea nuevo record %s" % recordnuevo)
                        elif RecordAcademico.objects.filter(status=True, inscripcion=inscripcion, asignaturamalla=equivalencia.asignaturamallasalto):
                            recordnuevo = RecordAcademico.objects.filter(status=True, inscripcion=inscripcion, asignaturamalla=equivalencia.asignaturamallasalto)[0]
                            recordnuevo.matriculas = recordantiguo.matriculas
                            recordnuevo.asignaturamalla = equivalencia.asignaturamallasalto
                            recordnuevo.asignatura = equivalencia.asignaturamallasalto.asignatura
                            recordnuevo.asignaturaold = recordantiguo.asignaturamalla.asignatura
                            recordnuevo.nota = recordantiguo.nota
                            recordnuevo.asistencia = recordantiguo.asistencia
                            recordnuevo.sinasistencia = recordantiguo.sinasistencia
                            recordnuevo.fecha = recordantiguo.fecha
                            recordnuevo.noaplica = recordantiguo.noaplica
                            recordnuevo.aprobada = recordantiguo.aprobada
                            recordnuevo.convalidacion = recordantiguo.convalidacion
                            recordnuevo.pendiente = recordantiguo.pendiente
                            recordnuevo.creditos = equivalencia.asignaturamallasalto.creditos
                            recordnuevo.horas = equivalencia.asignaturamallasalto.horas
                            recordnuevo.valida = recordantiguo.valida
                            recordnuevo.validapromedio = recordantiguo.validapromedio
                            recordnuevo.observaciones = observaciones
                            recordnuevo.homologada = homologada
                            recordnuevo.materiaregular = recordantiguo.materiaregular
                            recordnuevo.materiacurso = None
                            recordnuevo.completonota = recordantiguo.completonota
                            recordnuevo.completoasistencia = recordantiguo.completoasistencia
                            recordnuevo.fechainicio = recordantiguo.fechainicio
                            recordnuevo.fechafin = recordantiguo.fechafin
                            recordnuevo.suficiencia = recordantiguo.suficiencia
                            recordnuevo.asignaturamallahistorico = recordantiguo.asignaturamalla
                            recordnuevo.reverso = False
                            recordnuevo.save()
                        if recordnuevo:
                            historicos = HistoricoRecordAcademico.objects.filter(status=True, recordacademico=recordantiguo).update(recordacademico=recordnuevo, creditos=recordnuevo.creditos, horas=recordnuevo.horas, homologada=recordnuevo.homologada)
                            respaldo = RespaldoRecordAcademico.objects.filter(status=True,recordacademicooriginal=recordantiguo)

                            if not respaldo.exists():
                                respaldorecord = RespaldoRecordAcademico(recordacademicooriginal=recordantiguo, recordacademiconuevo=recordnuevo)
                                respaldorecord.save()
                            else:
                                respaldorecord = respaldo[0]
                                respaldorecord.recordacademiconuevo = recordnuevo
                                respaldorecord.save()
                            print(u"Record actualizado %s" % recordnuevo)
                    else:
                        hojadestino.write(fila, 3, "NO ENCONTRO RECORD ANTIGUO %s" % equivalencia.asignaturamalla)
                        fila += 1
                fila += 1
                time.sleep(3)
            lin += 1
        libdestino.save(output_folder + libre_origen)
        print(output_folder + libre_origen)
        print("Proceso finalizado. . .")


    except Exception as ex:
        transaction.set_rollback(True)
        print('error: %s' % ex)
        hojadestino.write(fila, 3, str(ex))
        fila += 1

def llenar_tabla_equivalencias_economia():
    try:
        miarchivo = openpyxl.load_workbook("matriz_equivalencia_economia.xlsx")
        lista = miarchivo.get_sheet_by_name('MALLA NUEVA')
        totallista = lista.rows
        a=0
        for filas in totallista:
            a += 1
            if a > 1 and f"{filas[4].value}".isdigit():
                idasignaturamallaanterior, idasignaturamallanueva = int(filas[4].value), int(filas[1].value)
                msg = ''
                if tablaeq := TablaEquivalenciaAsignaturas.objects.filter(asignaturamallasalto_id=idasignaturamallanueva, asignaturamalla_id=idasignaturamallaanterior).first():
                    tablaeq.status = True
                    msg = u"ACTUALIZA EQUIVALENCIA %s"
                else:
                    tablaeq = TablaEquivalenciaAsignaturas(asignaturamalla_id=idasignaturamallaanterior, asignaturamallasalto_id=idasignaturamallanueva)
                    msg = u"INSERTA EQUIVALENCIA %s"

                tablaeq.save()
                print(msg % f'{tablaeq.asignaturamallasalto} - {tablaeq.asignaturamalla}')

                print(u"Fila %s"%a)
        homologacion_economia()
    except Exception as ex:
            print('error: %s' % ex)


def llenar_tabla_equivalencias_comunicacion():
    try:
        miarchivo = openpyxl.load_workbook("matriz_equivalencia_comunicacion.xlsx")
        lista = miarchivo.get_sheet_by_name('MALLA_NUEVA')
        totallista = lista.rows
        a=0
        for filas in totallista:
            a += 1
            if a > 1  and f"{filas[1].value}".isdigit():
                idasignaturamallaanterior = int(filas[4].value)
                idasignaturamallanueva=int(filas[1].value)
                if not TablaEquivalenciaAsignaturas.objects.filter(status=True,asignaturamalla_id=idasignaturamallaanterior).exists():
                    tablaeq=TablaEquivalenciaAsignaturas(asignaturamalla_id=idasignaturamallaanterior,
                                                 asignaturamallasalto_id=idasignaturamallanueva)
                    tablaeq.save()
                    print(u"INSERTA EQUIVALENCIA %s"%tablaeq)
                else:
                    tablaeq=TablaEquivalenciaAsignaturas.objects.filter(status=True, asignaturamalla_id=idasignaturamallaanterior)[0]
                    tablaeq.asignaturamallasalto_id=idasignaturamallanueva
                    tablaeq.save()
                    print(u"ACTUALIZA EQUIVALENCIA %s" % tablaeq)
                print(u"Fila %s"%a)
        # homologacion()
    except Exception as ex:
            print('error: %s' % ex)

def homologacion_comunicacion():

    #verificar que todos los estudiantes tengan la misma malla
    cadena = ''
    linea, excluidos, conexito = 0, 0, 0
    try:
        libre_origen = '/homologacion_comunicacion.xls'
        fuentecabecera = easyxf(
            'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
        fuentenormal = easyxf(
            'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
        output_folder = MEDIA_ROOT
        output_folder = os.path.join(os.path.join(BASE_DIR))
        # liborigen = xlrd.open_workbook(output_folder + libre_origen)
        libdestino = xlwt.Workbook()
        hojadestino = libdestino.add_sheet('HOJA1')
        fil = 0
        columnas = [(u"CEDULA", 6000),
                   (u"APELLIDOS Y NOMBRES", 6000),
                   (u"OBSERVACIÓN", 6000)
        ]
        for col_num in range(len(columnas)):
            hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
            hojadestino.col(col_num).width = columnas[col_num][1]
        fila = 1
        titulacion = 0

        lin = 0
        miarchivo = openpyxl.load_workbook("CEDULA_COMUNICACION_final.xlsx")
        #miarchivo = openpyxl.load_workbook("prueba_2.xlsx")

        ws = miarchivo.get_sheet_by_name("Hoja1")
        worksheet = ws
        c = 0
        cont = 0
        periodo_id=177
        carrera_id=131
        mallaantigua_id=205
        mallanueva_id=488

        for row in worksheet.iter_rows(min_row=0):
            if lin >= 0:
                currentValues, cadena = [], ''
                for cell in row:
                    cadena += str(cell.value) + ' '
                    currentValues.append(str(cell.value))
                identificacion = currentValues[0]

                matricula = Matricula.objects.filter(status=True, inscripcion__carrera__id=carrera_id, inscripcion__persona__cedula=identificacion).first()
                cont += 1
                matricula.pasoayuda = True
                matricula.save()
                print(u"%s - %s" % (matricula, cont))
                inscripcion = matricula.inscripcion
                hojadestino.write(fila, 0, matricula.inscripcion.persona.identificacion(), fuentenormal)
                hojadestino.write(fila, 1, matricula.inscripcion.persona.nombre_completo(), fuentenormal)
                hojadestino.write(fila, 2, matricula.nivel.periodo.nombre, fuentenormal)

                if InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallaantigua_id).exists():
                    imantigua = InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallaantigua_id)[0]
                    imantigua.status = False
                    imantigua.save()
                    print(u"Desactiva antigua inscripcion -----------------------------")

                if not InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallanueva_id).exists():
                    imnueva = InscripcionMalla(inscripcion=inscripcion, malla_id=mallanueva_id)
                    imnueva.save()
                    print(u"Crea nueva inscripcion -----------------------------")

                equivalencias = TablaEquivalenciaAsignaturas.objects.filter(status=True, asignaturamalla__malla_id=mallaantigua_id).order_by('asignaturamallasalto__nivelmalla__orden')
                for equivalencia in equivalencias:
                    print(u"nueva - %s" % equivalencia.asignaturamallasalto)
                    recordantiguo = inscripcion.recordacademico_set.filter(status=True,asignaturamalla_id=equivalencia.asignaturamalla_id).first()

                    if recordantiguo:
                        print(u"anterior - %s" % equivalencia.asignaturamalla)
                        print(u"Record antiguo: %s" % recordantiguo)
                        recordnuevo = None
                        recordantiguo.status = False
                        recordantiguo.save(update_asignaturamalla=False)

                        if equivalencia.asignaturamallasalto_id in [10850,10853,10854,10859,10865]:
                            observaciones = recordantiguo.observaciones + " Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15"
                            homologada = True
                        else:
                            observaciones = recordantiguo.observaciones + " Migración con base a RESOLUCIÓN CGA-SO-9-2023-NO15 Y RESOLUCIÓN OCS-SO-18-2023-NO14"
                            homologada = False
                        if not RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                              asignaturamalla=equivalencia.asignaturamallasalto).exists():



                            recordnuevo = RecordAcademico(inscripcion=inscripcion,
                                                          matriculas=recordantiguo.matriculas,
                                                          asignaturamalla=equivalencia.asignaturamallasalto,
                                                          asignatura=equivalencia.asignaturamallasalto.asignatura,
                                                          asignaturaold_id=recordantiguo.asignatura.id,
                                                          nota=recordantiguo.nota,
                                                          asistencia=recordantiguo.asistencia,
                                                          sinasistencia=recordantiguo.sinasistencia,
                                                          fecha=recordantiguo.fecha,
                                                          noaplica=recordantiguo.noaplica,
                                                          aprobada=recordantiguo.aprobada,
                                                          convalidacion=recordantiguo.convalidacion,
                                                          pendiente=recordantiguo.pendiente,
                                                          creditos=equivalencia.asignaturamallasalto.creditos,
                                                          horas=equivalencia.asignaturamallasalto.horas,
                                                          valida=recordantiguo.valida,
                                                          validapromedio=recordantiguo.validapromedio,
                                                          observaciones=observaciones,
                                                          homologada=homologada,
                                                          materiaregular=recordantiguo.materiaregular,
                                                          materiacurso=None,
                                                          completonota=recordantiguo.completonota,
                                                          completoasistencia=recordantiguo.completoasistencia,
                                                          fechainicio=recordantiguo.fechainicio,
                                                          fechafin=recordantiguo.fechafin,
                                                          suficiencia=recordantiguo.suficiencia,
                                                          asignaturamallahistorico_id=recordantiguo.asignaturamalla.id,
                                                          reverso=False)
                            recordnuevo.save()
                            print(u"Crea nuevo record %s" % recordnuevo)


                        elif RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                            asignaturamalla=equivalencia.asignaturamallasalto):
                            recordnuevo = RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                         asignaturamalla=equivalencia.asignaturamallasalto)[0]
                            recordnuevo.matriculas = recordantiguo.matriculas
                            recordnuevo.asignaturamalla = equivalencia.asignaturamallasalto
                            recordnuevo.asignatura = equivalencia.asignaturamallasalto.asignatura
                            recordnuevo.asignaturaold = recordantiguo.asignaturamalla.asignatura
                            recordnuevo.nota = recordantiguo.nota
                            recordnuevo.asistencia = recordantiguo.asistencia
                            recordnuevo.sinasistencia = recordantiguo.sinasistencia
                            recordnuevo.fecha = recordantiguo.fecha
                            recordnuevo.noaplica = recordantiguo.noaplica
                            recordnuevo.aprobada = recordantiguo.aprobada
                            recordnuevo.convalidacion = recordantiguo.convalidacion
                            recordnuevo.pendiente = recordantiguo.pendiente
                            recordnuevo.creditos = equivalencia.asignaturamallasalto.creditos
                            recordnuevo.horas = equivalencia.asignaturamallasalto.horas
                            recordnuevo.valida = recordantiguo.valida
                            recordnuevo.validapromedio = recordantiguo.validapromedio
                            recordnuevo.observaciones = observaciones
                            recordnuevo.homologada = homologada
                            recordnuevo.materiaregular = recordantiguo.materiaregular
                            recordnuevo.materiacurso = None
                            recordnuevo.completonota = recordantiguo.completonota
                            recordnuevo.completoasistencia = recordantiguo.completoasistencia
                            recordnuevo.fechainicio = recordantiguo.fechainicio
                            recordnuevo.fechafin = recordantiguo.fechafin
                            recordnuevo.suficiencia = recordantiguo.suficiencia
                            recordnuevo.asignaturamallahistorico = recordantiguo.asignaturamalla
                            recordnuevo.reverso = False
                            recordnuevo.save()

                        if recordnuevo:
                            historicos = HistoricoRecordAcademico.objects.filter(status=True,
                                                                                 recordacademico=recordantiguo).update(recordacademico=recordnuevo,
                                                                                                                       creditos=recordnuevo.creditos,
                                                                                                                       horas=recordnuevo.horas,
                                                                                                                       homologada=recordnuevo.homologada)
                            respaldo = RespaldoRecordAcademico.objects.filter(status=True,recordacademicooriginal=recordantiguo)

                            if not respaldo.exists():
                                respaldorecord = RespaldoRecordAcademico(
                                    recordacademicooriginal=recordantiguo,
                                    recordacademiconuevo=recordnuevo
                                )
                                respaldorecord.save()
                            else:
                                respaldorecord = respaldo[0]
                                respaldorecord.recordacademiconuevo = recordnuevo
                                respaldorecord.save()
                            print(u"Record actualizado %s" % recordnuevo)


                    else:
                        hojadestino.write(fila, 3, "NO ENCONTRO RECORD ANTIGUO %s" % equivalencia.asignaturamalla)
                        fila += 1

                fila += 1

                time.sleep(3)

            lin += 1

        libdestino.save(output_folder + libre_origen)
        print(output_folder + libre_origen)
        print("Proceso finalizado. . .")


    except Exception as ex:
        transaction.set_rollback(True)
        print('error: %s' % ex)
        hojadestino.write(fila, 3, str(ex))
        fila += 1

def cambio_tutor_practicas1():
    try:
        miarchivo = openpyxl.load_workbook("practicas.xlsx")
        lista = miarchivo.get_sheet_by_name('Hoja2')
        totallista = lista.rows
        a=0
        idnuevotutor=4540
        for filas in totallista:
            a += 1
            if a > 1:
                idpractica = int(filas[0].value)
                practica = PracticasPreprofesionalesInscripcion.objects.get(pk=idpractica)
                practica.tutorunemi_id = idnuevotutor
                practica.save()
                profesor1 = Profesor.objects.get(pk=idnuevotutor)
                email = profesor1.persona.lista_emails_envio()
                profesor = profesor1.persona.nombre_completo_inverso()
                estudiante = practica.inscripcion.persona.nombre_completo_inverso()
                emailestudiante = practica.inscripcion.persona.lista_emails_envio()
                carrera = practica.inscripcion.carrera
                asunto = u"ASIGNACIÓN TUTOR ACADÉMICO"
                send_html_mail(asunto, "emails/tutor_practicas.html",
                               {'sistema': 'SGA',
                                'profesor': profesor,
                                'estudiante': estudiante,
                                'carrera': carrera},
                               email
                               , [], cuenta=CUENTAS_CORREOS[4][1])
                send_html_mail(asunto, "emails/tutor_practicas_alumno.html",
                               {'sistema': 'SGA',
                                'profesor': profesor,
                                'estudiante': estudiante}, emailestudiante
                               , [], cuenta=CUENTAS_CORREOS[4][1])

                idprof = idnuevotutor
                profesor1 = Profesor.objects.get(pk=idprof)
                asunto = u"ASIGNACIÓN TUTOR ACADÉMICO PRÁCTICAS PREPROFESIONALES"
                para = profesor1.persona
                observacion = 'Se le comunica que ha sido designado como tutor académico a (el/la) estudiante: {} de la carrera: {}'.format(
                    estudiante, carrera)
                notificacion2(asunto, observacion, para, None, '/pro_cronograma?action=listatutorias',
                             practica.pk, 1, 'sga',
                             DetallePreInscripcionPracticasPP)



    except Exception as ex:
            print('error: %s' % ex)


def cambio_tutor_practicas2():
    try:
        miarchivo = openpyxl.load_workbook("practicas.xlsx")
        lista = miarchivo.get_sheet_by_name('Hoja3')
        totallista = lista.rows
        a=0
        idnuevotutor=4312
        for filas in totallista:
            a += 1
            if a > 1:
                idpractica = int(filas[0].value)
                practica = PracticasPreprofesionalesInscripcion.objects.get(pk=idpractica)
                practica.tutorunemi_id = idnuevotutor
                practica.save()
                profesor1 = Profesor.objects.get(pk=idnuevotutor)
                email = profesor1.persona.lista_emails_envio()
                profesor = profesor1.persona.nombre_completo_inverso()
                estudiante = practica.inscripcion.persona.nombre_completo_inverso()
                emailestudiante = practica.inscripcion.persona.lista_emails_envio()
                carrera = practica.inscripcion.carrera
                asunto = u"ASIGNACIÓN TUTOR ACADÉMICO"
                send_html_mail(asunto, "emails/tutor_practicas.html",
                               {'sistema': 'SGA',
                                'profesor': profesor,
                                'estudiante': estudiante,
                                'carrera': carrera},
                               email
                               , [], cuenta=CUENTAS_CORREOS[4][1])
                send_html_mail(asunto, "emails/tutor_practicas_alumno.html",
                               {'sistema': 'SGA',
                                'profesor': profesor,
                                'estudiante': estudiante}, emailestudiante
                               , [], cuenta=CUENTAS_CORREOS[4][1])

                idprof = idnuevotutor
                profesor1 = Profesor.objects.get(pk=idprof)
                asunto = u"ASIGNACIÓN TUTOR ACADÉMICO PRÁCTICAS PREPROFESIONALES"
                para = profesor1.persona
                observacion = 'Se le comunica que ha sido designado como tutor académico a (el/la) estudiante: {} de la carrera: {}'.format(
                    estudiante, carrera)
                notificacion2(asunto, observacion, para, None, '/pro_cronograma?action=listatutorias',
                             practica.pk, 1, 'sga',
                             DetallePreInscripcionPracticasPP)



    except Exception as ex:
            print('error: %s' % ex)



def cambio_supervisor_practicas():
    try:
        miarchivo = openpyxl.load_workbook("practicas.xlsx")
        lista = miarchivo.get_sheet_by_name('Hoja4')
        totallista = lista.rows
        a=0
        idnuevotutor=4540
        for filas in totallista:
            a += 1
            if a > 1:
                idpractica = int(filas[0].value)
                practica = PracticasPreprofesionalesInscripcion.objects.get(pk=idpractica)
                practica.supervisor_id = idnuevotutor
                practica.save()
                #profesor1 = Profesor.objects.get(pk=idnuevotutor)
                # email = profesor1.persona.lista_emails_envio()
                # profesor = profesor1.persona.nombre_completo_inverso()
                # estudiante = practica.inscripcion.persona.nombre_completo_inverso()
                # emailestudiante = practica.inscripcion.persona.lista_emails_envio()
                # carrera = practica.inscripcion.carrera


                # asunto = u"ASIGNACIÓN SUPERVISOR ACADÉMICO"
                # send_html_mail(asunto, "emails/supervisor_practicas.html",
                #                {'sistema': 'SGA',
                #                 'profesor': profesor,
                #                 'estudiante': estudiante,
                #                 'carrera': carrera}, email
                #                , [], cuenta=CUENTAS_CORREOS[4][1])
                # send_html_mail(asunto, "emails/supervisor_practicas_alumno.html",
                #                {'sistema': 'SGA',
                #                 'profesor': profesor,
                #                 'estudiante': estudiante}, emailestudiante
                #                , [], cuenta=CUENTAS_CORREOS[4][1])
                # idprof = idnuevotutor
                # profesor1 = Profesor.objects.get(pk=idprof)
                # asunto = u"ASIGNACIÓN SUPERVISOR ACADÉMICO PRÁCTICAS PREPROFESIONALES"
                # para = profesor1.persona
                # observacion = 'Se le comunica que ha sido designado como supervisor académico a (el/la) estudiante: {} de la carrera: {}'.format(
                #     estudiante, carrera)
                # notificacion2(asunto, observacion, para, None, '/pro_cronograma?action=listatutorias',
                #               practica.pk, 1, 'sga',
                #               DetallePreInscripcionPracticasPP)
                #
        print('FIN PROCESO------------')



    except Exception as ex:
            print('error: %s' % ex)

def creacion_examen_segundo_parcial_en_linea__recuperacion(ePeriodo):
    from Moodle_Funciones import CrearExamenMoodle
    ePersona = Persona.objects.get(pk=29898)
    eNotificacion = Notificacion(cuerpo='Creación examenes 2S 2023 PREGRADO',
                                 titulo=f'(En proceso) Creación de examenes 2S 2023 PREGRADO',
                                 destinatario=ePersona,
                                 url='',
                                 prioridad=1,
                                 app_label='SGA',
                                 fecha_hora_visible=datetime.now() + timedelta(days=5),
                                 tipo=2,
                                 en_proceso=True)
    eNotificacion.save()

    eNiveles = Nivel.objects.filter(status=True, periodo=ePeriodo, nivellibrecoordinacion__coordinacion_id__in=[1, 2, 3, 4, 5])
    eMaterias = Materia.objects.filter(status=True, nivel__in=eNiveles,  modeloevaluativo_id=7, asignaturamalla__malla__modalidad_id__in=[3])
    total = eMaterias.count()
    contador = 0
    for eMateria in eMaterias:
        eProfesor = eMateria.profesor_principal()
        test = TestSilaboSemanal.objects.filter(status=True, detallemodelo_id=39, silabosemanal__silabo__materia=eMateria).exists()
        if not test:
            with transaction.atomic():
                try:
                    contador += 1
                    print(f"({contador}/{total}) Materia: ", eMateria.__str__())
                    ePlanificacionClaseSilabos = PlanificacionClaseSilabo.objects.filter(status=True, parcial=2, examen=True, tipoplanificacion__periodo=ePeriodo, tipoplanificacion__planificacionclasesilabo_materia__materia=eMateria)
                    eSilabo = Silabo.objects.filter(materia=eMateria, codigoqr=True, status=True).first()
                    if not eSilabo:
                        raise NameError(u"No se encontro silabo de la materia: %s" % eMateria.__str__())
                    if (ePlanificacion := ePlanificacionClaseSilabos.first()) is not None:
                        numsemana = ePlanificacion.semana
                        try:
                            eSilaboSemanal = SilaboSemanal.objects.get(silabo=eSilabo, numsemana=numsemana , status=True)
                        except ObjectDoesNotExist:
                            eSilaboSemanal = SilaboSemanal(silabo=eSilabo,
                                                          numsemana=numsemana,
                                                          fechainiciosemana=ePlanificacion.fechainicio,
                                                          fechafinciosemana=ePlanificacion.fechafin,
                                                          examen=True,
                                                          semana=ePlanificacion.fechainicio.isocalendar()[1])
                        eSilaboSemanal.fechainiciosemana=ePlanificacion.fechainicio
                        eSilaboSemanal.fechafinciosemana=ePlanificacion.fechafin
                        eSilaboSemanal.semana=ePlanificacion.fechainicio.isocalendar()[1]
                        eSilaboSemanal.examen = True
                        eSilaboSemanal.save()
                        eDetalleModeloEvaluativo = DetalleModeloEvaluativo.objects.filter(nombre="RE", modelo=eMateria.modeloevaluativo, status=True).first()
                        try:
                            eTestSilaboSemanal = TestSilaboSemanal.objects.get(silabosemanal=eSilaboSemanal, detallemodelo=eDetalleModeloEvaluativo, tiporecurso_id=11)
                        except ObjectDoesNotExist:
                            eTestSilaboSemanal = TestSilaboSemanal(silabosemanal=eSilaboSemanal,
                                                                   detallemodelo=eDetalleModeloEvaluativo,
                                                                   tiporecurso_id=11)
                        instruccion = """- Para resolver el presente Examen, dispone de un solo intento. 
    - Acceda al examen en la hora y fecha programada.
    - No se puede prorrogar el horario del examen. 
    - Coloque la clave de acceso al examen en caso que este lo solicite.
    - Revise el tiempo asignado para su examen, el examen debe ser resuelto en el tiempo máximo indicado, si comienza el intento tiempo después, ese tiempo se le descontará de la hora preestablecida.
    - Revise el número de preguntas a desarrollar.
    - Asegúrese que su promedio final previo a este examen sea mayor a 39,5 y menor a 69,5 puntos.
    - Lea atentamente cada pregunta antes de registrar su respuesta.
    - Verifique la modalidad de navegación del examen (LIBRE o SECUENCIAL).
    - El examen está estructurado con todo el contenido y materiales del curso desde la primera semana, recuerde que el examen de recuperación es acumulativo.
    - Cualquier acto de deshonestidad académica será considerado como una falta y motivo de la inmediata suspensión del examen.
    - El examen tendrá una puntuación máxima de 100 puntos. 
    - Una vez finalizado el examen no olvide dar clic en terminar el intento.
    """
                        recomendacion = """Lea detenidamente las preguntas y responda."""
                        eTestSilaboSemanal.estado_id=2
                        eTestSilaboSemanal.calificar=True
                        eTestSilaboSemanal.nombretest='EXAMEN_RE'
                        eTestSilaboSemanal.instruccion=instruccion
                        eTestSilaboSemanal.recomendacion=recomendacion
                        eTestSilaboSemanal.fechadesde=ePlanificacion.fechainicio
                        eTestSilaboSemanal.horadesde=datetime(ePlanificacion.fechainicio.year, ePlanificacion.fechainicio.month, ePlanificacion.fechainicio.day, 0, 1)
                        eTestSilaboSemanal.fechahasta=ePlanificacion.fechafin
                        eTestSilaboSemanal.horahasta=datetime(ePlanificacion.fechafin.year, ePlanificacion.fechafin.month, ePlanificacion.fechafin.day, 23, 59)
                        eTestSilaboSemanal.vecesintento=1
                        eTestSilaboSemanal.navegacion=2
                        eTestSilaboSemanal.tiempoduracion=60
                        eTestSilaboSemanal.password='EX-RE-24'
                        eTestSilaboSemanal.save()
                        if eProfesor:
                            value, msg = CrearExamenMoodle(eTestSilaboSemanal.id, eProfesor.persona)
                            if not value:
                                raise NameError(msg)
                            eMateria.actualizarhtml = True
                            eMateria.save()
                        print(f"({contador}/{total}) Test: {eTestSilaboSemanal.__str__()} - Materia: ", eMateria.__str__())
                except Exception as ex:
                    transaction.set_rollback(True)
                    msg = ex.__str__()
                    print(msg)

    eNotificacion.url = None
    eNotificacion.titulo = f'(Finalizado) Creación de examenes 2S 2023 PREGRADO'
    eNotificacion.save()
    print("Proceso finalizado . . .")
# creacion_examen_segundo_parcial_en_linea__recuperacion(224)




def creacion_examen_segundo_parcial_presencial__recuperacion(ePeriodo):
    #from Moodle_Funciones import CrearExamenMoodle
    ePersona = Persona.objects.get(pk=29898)
    eNotificacion = Notificacion(cuerpo='Creación examenes 2S 2023 PREGRADO',
                                 titulo=f'(En proceso) Creación de examenes 2S 2023 PREGRADO',
                                 destinatario=ePersona,
                                 url='',
                                 prioridad=1,
                                 app_label='SGA',
                                 fecha_hora_visible=datetime.now() + timedelta(days=5),
                                 tipo=2,
                                 en_proceso=True)
    eNotificacion.save()

    eNiveles = Nivel.objects.filter(status=True, periodo=ePeriodo, nivellibrecoordinacion__coordinacion_id__in=[1, 2, 3, 4, 5])
    eMaterias = Materia.objects.filter(pk=68377, status=True, nivel__in=eNiveles,  modeloevaluativo_id=7, asignaturamalla__malla__modalidad_id__in=[1, 2])
    total = eMaterias.count()
    contador = 0
    for eMateria in eMaterias:
        eProfesor = eMateria.profesor_principal()
        test = TestSilaboSemanal.objects.filter(status=True, detallemodelo_id=39, silabosemanal__silabo__materia=eMateria).exists()
        if not test:
            with transaction.atomic():
                try:
                    contador += 1
                    print(f"({contador}/{total}) Materia: ", eMateria.__str__())
                    ePlanificacionClaseSilabos = PlanificacionClaseSilabo.objects.filter(status=True, parcial=2, examen=True, tipoplanificacion__periodo=ePeriodo, tipoplanificacion__planificacionclasesilabo_materia__materia=eMateria)
                    eSilabo = Silabo.objects.filter(materia=eMateria, codigoqr=True, status=True).first()
                    if not eSilabo:
                        raise NameError(u"No se encontro silabo de la materia: %s" % eMateria.__str__())
                    if (ePlanificacion := ePlanificacionClaseSilabos.first()) is not None:
                        numsemana = ePlanificacion.semana
                        try:
                            eSilaboSemanal = SilaboSemanal.objects.get(silabo=eSilabo, numsemana=numsemana , status=True)
                        except ObjectDoesNotExist:
                            eSilaboSemanal = SilaboSemanal(silabo=eSilabo,
                                                          numsemana=numsemana,
                                                          fechainiciosemana=ePlanificacion.fechainicio,
                                                          fechafinciosemana=ePlanificacion.fechafin,
                                                          examen=True,
                                                          semana=ePlanificacion.fechainicio.isocalendar()[1])
                        eSilaboSemanal.fechainiciosemana=ePlanificacion.fechainicio
                        eSilaboSemanal.fechafinciosemana=ePlanificacion.fechafin
                        eSilaboSemanal.semana=ePlanificacion.fechainicio.isocalendar()[1]
                        eSilaboSemanal.examen = True
                        eSilaboSemanal.save()
                        eDetalleModeloEvaluativo = DetalleModeloEvaluativo.objects.filter(nombre="RE", modelo=eMateria.modeloevaluativo, status=True).first()
                        try:
                            eTestSilaboSemanal = TestSilaboSemanal.objects.get(silabosemanal=eSilaboSemanal, detallemodelo=eDetalleModeloEvaluativo, tiporecurso_id=11)
                        except ObjectDoesNotExist:
                            eTestSilaboSemanal = TestSilaboSemanal(silabosemanal=eSilaboSemanal,
                                                                   detallemodelo=eDetalleModeloEvaluativo,
                                                                   tiporecurso_id=11)
                        instruccion = """- Para resolver el presente Examen, dispone de un solo intento. 
    - Acceda al examen en la hora y fecha programada.
    - No se puede prorrogar el horario del examen. 
    - Coloque la clave de acceso al examen en caso que este lo solicite.
    - Revise el tiempo asignado para su examen, el examen debe ser resuelto en el tiempo máximo indicado, si comienza el intento tiempo después, ese tiempo se le descontará de la hora preestablecida.
    - Revise el número de preguntas a desarrollar.
    - Asegúrese que su promedio final previo a este examen sea mayor a 39,5 y menor a 69,5 puntos.
    - Lea atentamente cada pregunta antes de registrar su respuesta.
    - Verifique la modalidad de navegación del examen (LIBRE o SECUENCIAL).
    - El examen está estructurado con todo el contenido y materiales del curso desde la primera semana, recuerde que el examen de recuperación es acumulativo.
    - Cualquier acto de deshonestidad académica será considerado como una falta y motivo de la inmediata suspensión del examen.
    - El examen tendrá una puntuación máxima de 100 puntos. 
    - Una vez finalizado el examen no olvide dar clic en terminar el intento.
    """
                        recomendacion = """Lea detenidamente las preguntas y responda."""
                        eTestSilaboSemanal.estado_id=2
                        eTestSilaboSemanal.calificar=True
                        eTestSilaboSemanal.nombretest='EXAMEN_RE'
                        eTestSilaboSemanal.instruccion=instruccion
                        eTestSilaboSemanal.recomendacion=recomendacion
                        eTestSilaboSemanal.fechadesde=ePlanificacion.fechainicio
                        eTestSilaboSemanal.horadesde=datetime(ePlanificacion.fechainicio.year, ePlanificacion.fechainicio.month, ePlanificacion.fechainicio.day, 0, 1)
                        eTestSilaboSemanal.fechahasta=ePlanificacion.fechafin
                        eTestSilaboSemanal.horahasta=datetime(ePlanificacion.fechafin.year, ePlanificacion.fechafin.month, ePlanificacion.fechafin.day, 23, 59)
                        eTestSilaboSemanal.vecesintento=1
                        eTestSilaboSemanal.navegacion=2
                        eTestSilaboSemanal.tiempoduracion=60
                        eTestSilaboSemanal.password='EX-RE-24'
                        eTestSilaboSemanal.save()
                        # if eProfesor:
                        #     value, msg = CrearExamenMoodle(eTestSilaboSemanal.id, eProfesor.persona)
                        #     if not value:
                        #         raise NameError(msg)
                        #     eMateria.actualizarhtml = True
                        #     eMateria.save()
                        # print(f"({contador}/{total}) Test: {eTestSilaboSemanal.__str__()} - Materia: ", eMateria.__str__())
                except Exception as ex:
                    transaction.set_rollback(True)
                    msg = ex.__str__()
                    print(msg)

    eNotificacion.url = None
    eNotificacion.titulo = f'(Finalizado) Creación de examenes 2S 2023 PREGRADO'
    eNotificacion.save()
    print("Proceso finalizado . . .")
# creacion_examen_segundo_parcial_presencial__recuperacion(224)


def homologacion_idiomas():
    #verificar que todos los estudiantes tengan la misma malla
    cadena = ''
    linea, excluidos, conexito = 0, 0, 0
    try:
        libre_origen = '/homologacion_idiomas_2_1.xls'
        fuentecabecera = easyxf(
            'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
        fuentenormal = easyxf(
            'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')

        output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
        # liborigen = xlrd.open_workbook(output_folder + libre_origen)
        libdestino = xlwt.Workbook()
        hojadestino = libdestino.add_sheet('Sheet1')
        fil = 0
        columnas = [(u"CEDULA", 6000),
                    (u"APELLIDOS Y NOMBRES", 6000),
                    (u"OBSERVACIÓN", 6000),
                    (u"HORAS PRACTICAS", 6000),
                    (u"HORAS VINCULACION", 6000),
                    (u"OBSERVACIÓN", 6000)
                    ]
        for col_num in range(len(columnas)):
            hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
            hojadestino.col(col_num).width = columnas[col_num][1]
        fila = 1
        titulacion = 0

        lin = 0
        #miarchivo = openpyxl.load_workbook("CEDULA_COMUNICACION_final.xlsx")
        miarchivo = openpyxl.load_workbook("ilinea.xlsx")
        #miarchivo = openpyxl.load_workbook("prueba_2.xlsx")

        ws = miarchivo.get_sheet_by_name("primero")
        worksheet = ws
        c = 0
        cont = 0
        periodo_id=224
        carrera_id=129
        mallaantigua_id=198
        mallanueva_id=492

        for row in worksheet.iter_rows(min_row=0):
            if lin >= 0:
                currentValues, cadena = [], ''
                for cell in row:
                    cadena += str(cell.value) + ' '
                    currentValues.append(str(cell.value))
                identificacion = currentValues[0]

                if not identificacion:
                    break

                matricula = Matricula.objects.filter(status=True, inscripcion__carrera__id=carrera_id, inscripcion__persona__cedula=identificacion).first()
                cont += 1
                matricula.pasoayuda = True
                matricula.save()
                print(u"%s - %s" % (matricula, cont))
                inscripcion = matricula.inscripcion
                hojadestino.write(fila, 0, matricula.inscripcion.persona.identificacion(), fuentenormal)
                hojadestino.write(fila, 1, matricula.inscripcion.persona.nombre_completo(), fuentenormal)
                hojadestino.write(fila, 2, matricula.nivel.periodo.nombre, fuentenormal)

                practicaspp = haber_cumplido_horas_creditos_practicas_preprofesionales(inscripcion.id)
                horasvinculacion = haber_cumplido_horas_creditos_vinculacion(inscripcion.id)

                if InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallaantigua_id).exists():
                    imantigua = InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallaantigua_id)[0]
                    imantigua.status = False
                    imantigua.save()
                    print(u"Desactiva antigua inscripcion -----------------------------")

                if not InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallanueva_id).exists():
                    imnueva = InscripcionMalla(inscripcion=inscripcion, malla_id=mallanueva_id)
                    imnueva.save()
                    print(u"Crea nueva inscripcion -----------------------------")

                equivalencias = TablaEquivalenciaAsignaturas.objects.filter(status=True, asignaturamalla__malla_id=mallaantigua_id).order_by('asignaturamallasalto__nivelmalla__orden')
                cont_asig_vinculacion_aprobadas = 0
                horasfalta = 0
                fechainicioitinerario = None
                fechafinitinerario = None
                for equivalencia in equivalencias:
                    print(u"nueva - %s" % equivalencia.asignaturamallasalto)
                    recordantiguo = inscripcion.recordacademico_set.filter(status=True,asignaturamalla_id=equivalencia.asignaturamalla_id).first()

                    if recordantiguo:
                        print(u"anterior - %s" % equivalencia.asignaturamalla)
                        print(u"Record antiguo: %s" % recordantiguo)
                        recordnuevo = None
                        recordantiguo.status = False
                        recordantiguo.save(update_asignaturamalla=False)

                        if equivalencia.asignaturamallasalto_id in [11018,11021,11020,10990,10991,11005,10993,10997,11007]:
                            observaciones = recordantiguo.observaciones + " Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15"
                            homologada = True


                        else:
                            observaciones = recordantiguo.observaciones + " Migración con base a RESOLUCIÓN CGA-SO-9-2023-NO15 Y RESOLUCIÓN OCS-SO-18-2023-NO14"
                            homologada = False
                        if not RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                              asignaturamalla=equivalencia.asignaturamallasalto).exists():



                            recordnuevo = RecordAcademico(inscripcion=inscripcion,
                                                          matriculas=recordantiguo.matriculas,
                                                          asignaturamalla=equivalencia.asignaturamallasalto,
                                                          asignatura=equivalencia.asignaturamallasalto.asignatura,
                                                          asignaturaold_id=recordantiguo.asignatura.id,
                                                          nota=recordantiguo.nota,
                                                          asistencia=recordantiguo.asistencia,
                                                          sinasistencia=recordantiguo.sinasistencia,
                                                          fecha=recordantiguo.fecha,
                                                          noaplica=recordantiguo.noaplica,
                                                          aprobada=recordantiguo.aprobada,
                                                          convalidacion=recordantiguo.convalidacion,
                                                          pendiente=recordantiguo.pendiente,
                                                          creditos=equivalencia.asignaturamallasalto.creditos,
                                                          horas=equivalencia.asignaturamallasalto.horas,
                                                          valida=recordantiguo.valida,
                                                          validapromedio=recordantiguo.validapromedio,
                                                          observaciones=observaciones,
                                                          homologada=homologada,
                                                          materiaregular=recordantiguo.materiaregular,
                                                          materiacurso=None,
                                                          completonota=recordantiguo.completonota,
                                                          completoasistencia=recordantiguo.completoasistencia,
                                                          fechainicio=recordantiguo.fechainicio,
                                                          fechafin=recordantiguo.fechafin,
                                                          suficiencia=recordantiguo.suficiencia,
                                                          asignaturamallahistorico_id=recordantiguo.asignaturamalla.id,
                                                          reverso=False)
                            recordnuevo.save()
                            print(u"Crea nuevo record %s" % recordnuevo)


                        elif RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                            asignaturamalla=equivalencia.asignaturamallasalto):
                            recordnuevo = RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                         asignaturamalla=equivalencia.asignaturamallasalto)[0]
                            recordnuevo.matriculas = recordantiguo.matriculas
                            recordnuevo.asignaturamalla = equivalencia.asignaturamallasalto
                            recordnuevo.asignatura = equivalencia.asignaturamallasalto.asignatura
                            recordnuevo.asignaturaold = recordantiguo.asignaturamalla.asignatura
                            recordnuevo.nota = recordantiguo.nota
                            recordnuevo.asistencia = recordantiguo.asistencia
                            recordnuevo.sinasistencia = recordantiguo.sinasistencia
                            recordnuevo.fecha = recordantiguo.fecha
                            recordnuevo.noaplica = recordantiguo.noaplica
                            recordnuevo.aprobada = recordantiguo.aprobada
                            recordnuevo.convalidacion = recordantiguo.convalidacion
                            recordnuevo.pendiente = recordantiguo.pendiente
                            recordnuevo.creditos = equivalencia.asignaturamallasalto.creditos
                            recordnuevo.horas = equivalencia.asignaturamallasalto.horas
                            recordnuevo.valida = recordantiguo.valida
                            recordnuevo.validapromedio = recordantiguo.validapromedio
                            recordnuevo.observaciones = observaciones
                            recordnuevo.homologada = homologada
                            recordnuevo.materiaregular = recordantiguo.materiaregular
                            recordnuevo.materiacurso = None
                            recordnuevo.completonota = recordantiguo.completonota
                            recordnuevo.completoasistencia = recordantiguo.completoasistencia
                            recordnuevo.fechainicio = recordantiguo.fechainicio
                            recordnuevo.fechafin = recordantiguo.fechafin
                            recordnuevo.suficiencia = recordantiguo.suficiencia
                            recordnuevo.asignaturamallahistorico = recordantiguo.asignaturamalla
                            recordnuevo.reverso = False
                            recordnuevo.save()

                        if recordnuevo:
                            historicos = HistoricoRecordAcademico.objects.filter(status=True,
                                                                                 recordacademico=recordantiguo).update(recordacademico=recordnuevo,
                                                                                                                       creditos=recordnuevo.creditos,
                                                                                                                       horas=recordnuevo.horas,
                                                                                                                       homologada=recordnuevo.homologada)
                            respaldo = RespaldoRecordAcademico.objects.filter(status=True,recordacademicooriginal=recordantiguo)



                            if equivalencia.asignaturamallasalto_id in [11018,11021,11020,10990,10991,11005,10993,10997,11007]:
                                if not practicaspp or not horasvinculacion:
                                    if recordnuevo.aprobada:
                                        profesor = None
                                        if recordnuevo.materiaregular:
                                            profesor = recordnuevo.materiaregular.profesor_principal()
                                        elif recordnuevo.materiacurso:
                                            profesor = recordnuevo.materiaregular.profesor()
                                        if equivalencia.asignaturamallasalto_id == 11018:

                                            itinerarioprimero = ItinerariosMalla.objects.get(status=True, malla_id=mallaantigua_id, nivel_id=1)
                                            iprimeronuevo = ItinerariosMalla.objects.get(status=True, malla_id=mallanueva_id, nivel_id=1)


                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerarioprimero).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerarioprimero).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerarioprimero).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=iprimeronuevo.horas_practicas,
                                                                                                         nivelmalla=iprimeronuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=iprimeronuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerarioprimero).update(
                                                itinerario=iprimeronuevo)

                                        if equivalencia.asignaturamallasalto_id == 11021:
                                            itinerariosegundo = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=2)
                                            isegundonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=2)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerariosegundo).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerariosegundo).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerariosegundo).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=isegundonuevo.horas_practicas,
                                                                                                         nivelmalla=isegundonuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=isegundonuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerariosegundo).update(
                                                itinerario=isegundonuevo)

                                        if equivalencia.asignaturamallasalto_id == 11020:
                                            itinerariotercero = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=3)
                                            iterceronuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=3)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerariotercero).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerariotercero).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerariotercero).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=iterceronuevo.horas_practicas,
                                                                                                         nivelmalla=iterceronuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=iterceronuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerariotercero).update(
                                                itinerario=iterceronuevo)

                                        if equivalencia.asignaturamallasalto_id == 10990:
                                            itinerariocuarto = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=4)
                                            icuartonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=4)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerariocuarto).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerariocuarto).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerariocuarto).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=icuartonuevo.horas_practicas,
                                                                                                         nivelmalla=icuartonuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=icuartonuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerariocuarto).update(
                                                itinerario=icuartonuevo)

                                        if equivalencia.asignaturamallasalto_id == 10991:
                                            itinerarioquinto = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=5)
                                            iquintonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=5)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerarioquinto).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerarioquinto).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerarioquinto).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=iquintonuevo.horas_practicas,
                                                                                                         nivelmalla=iquintonuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=iquintonuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerarioquinto).update(
                                                itinerario=iquintonuevo)

                                        if equivalencia.asignaturamallasalto_id == 11005:
                                            itinerariosexto = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=6)
                                            isextonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=6)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerariosexto).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerariosexto).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerariosexto).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=isextonuevo.horas_practicas,
                                                                                                         nivelmalla=isextonuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=isextonuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerariosexto).update(
                                                itinerario=isextonuevo)

                                        if equivalencia.asignaturamallasalto_id == 10993:
                                            ####################################### VINCULACION ###################################################
                                            if not horasvinculacion:
                                                if inscripcion.numero_horas_proyectos_vinculacion() < 80:
                                                    horasfalta = 80 - inscripcion.numero_horas_proyectos_vinculacion()
                                                    vinculacion = ParticipantesMatrices(status=True,
                                                                                        matrizevidencia_id=2,
                                                                                        proyecto_id=601,
                                                                                        inscripcion=inscripcion,
                                                                                        horas=horasfalta,
                                                                                        registrohorasdesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                        registrohorashasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                        estado=1
                                                                                        )
                                                    vinculacion.save()
                                            ######################################################################################################
                                            itinerarioseptimo = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=7)
                                            iseptimonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=7)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerarioseptimo).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerarioseptimo).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerarioseptimo).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=iseptimonuevo.horas_practicas,
                                                                                                         nivelmalla=iseptimonuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=iseptimonuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerarioseptimo).update(
                                                itinerario=iseptimonuevo)

                                        if equivalencia.asignaturamallasalto_id == 10997:
                                            itinerariooctavo = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=8)
                                            ioctavonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=8)
                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerariooctavo).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerariooctavo).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerariooctavo).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=ioctavonuevo.horas_practicas,
                                                                                                         nivelmalla=ioctavonuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=ioctavonuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerariooctavo).update(
                                                itinerario=ioctavonuevo)

                                        if equivalencia.asignaturamallasalto_id == 11007:
                                            ####################################### VINCULACION ###################################################
                                            if not horasvinculacion:
                                                if inscripcion.numero_horas_proyectos_vinculacion() < 160:
                                                    horasfalta = 160 - inscripcion.numero_horas_proyectos_vinculacion()
                                                    vinculacion = ParticipantesMatrices(status=True,
                                                                                        matrizevidencia_id=2,
                                                                                        proyecto_id=601,
                                                                                        inscripcion=inscripcion,
                                                                                        horas=horasfalta,
                                                                                        registrohorasdesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                        registrohorashasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                        estado=1
                                                                                        )
                                                    vinculacion.save()
                                            ######################################################################################################
                                            itinerarionoveno = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=9)
                                            inovenonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=9)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerarionoveno).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerarionoveno).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerarionoveno).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=inovenonuevo.horas_practicas,
                                                                                                         nivelmalla=inovenonuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=inovenonuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerarionoveno).update(
                                                itinerario=inovenonuevo)


                            if not respaldo.exists():
                                respaldorecord = RespaldoRecordAcademico(
                                    recordacademicooriginal=recordantiguo,
                                    recordacademiconuevo=recordnuevo
                                )
                                respaldorecord.save()
                            else:
                                respaldorecord = respaldo[0]
                                respaldorecord.recordacademiconuevo = recordnuevo
                                respaldorecord.save()
                            print(u"Record actualizado %s" % recordnuevo)

                    else:
                        hojadestino.write(fila, 3, "NO ENCONTRO RECORD ANTIGUO %s" % equivalencia.asignaturamalla)
                        fila += 1

                practicasppf = inscripcion.numero_horas_practicas_pre_profesionales()
                hojadestino.write(fila, 3, practicasppf, fuentenormal)
                horasvinculacionf = inscripcion.numero_horas_proyectos_vinculacion()
                hojadestino.write(fila, 4, horasvinculacionf, fuentenormal)

                fila += 1

                time.sleep(1)

            lin += 1

        libdestino.save(output_folder + libre_origen)
        print(output_folder + libre_origen)
        print("Proceso finalizado. . .")

    except Exception as ex:
        transaction.set_rollback(True)
        print('error: %s' % ex)
        hojadestino.write(fila, 3, str(ex))
        fila += 1

def homologacion_idiomas2():
    #verificar que todos los estudiantes tengan la misma malla
    cadena = ''
    linea, excluidos, conexito = 0, 0, 0
    try:
        libre_origen = '/homologacion_idiomas_2_2.xls'
        fuentecabecera = easyxf(
            'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
        fuentenormal = easyxf(
            'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')

        output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
        # liborigen = xlrd.open_workbook(output_folder + libre_origen)
        libdestino = xlwt.Workbook()
        hojadestino = libdestino.add_sheet('Sheet1')
        fil = 0
        columnas = [(u"CEDULA", 6000),
                    (u"APELLIDOS Y NOMBRES", 6000),
                    (u"OBSERVACIÓN", 6000),
                    (u"HORAS PRACTICAS", 6000),
                    (u"HORAS VINCULACION", 6000),
                    (u"OBSERVACIÓN", 6000)
                    ]
        for col_num in range(len(columnas)):
            hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
            hojadestino.col(col_num).width = columnas[col_num][1]
        fila = 1
        titulacion = 0

        lin = 0
        #miarchivo = openpyxl.load_workbook("CEDULA_COMUNICACION_final.xlsx")
        miarchivo = openpyxl.load_workbook("ilinea.xlsx")
        #miarchivo = openpyxl.load_workbook("prueba_2.xlsx")

        ws = miarchivo.get_sheet_by_name("segundo")
        worksheet = ws
        c = 0
        cont = 0
        periodo_id=224
        carrera_id=129
        mallaantigua_id=198
        mallanueva_id=492

        for row in worksheet.iter_rows(min_row=0):
            if lin >= 0:
                currentValues, cadena = [], ''
                for cell in row:
                    cadena += str(cell.value) + ' '
                    currentValues.append(str(cell.value))
                identificacion = currentValues[0]

                if not identificacion:
                    break

                matricula = Matricula.objects.filter(status=True, inscripcion__carrera__id=carrera_id, inscripcion__persona__cedula=identificacion).first()
                cont += 1
                matricula.pasoayuda = True
                matricula.save()
                print(u"%s - %s" % (matricula, cont))
                inscripcion = matricula.inscripcion
                hojadestino.write(fila, 0, matricula.inscripcion.persona.identificacion(), fuentenormal)
                hojadestino.write(fila, 1, matricula.inscripcion.persona.nombre_completo(), fuentenormal)
                hojadestino.write(fila, 2, matricula.nivel.periodo.nombre, fuentenormal)

                practicaspp = haber_cumplido_horas_creditos_practicas_preprofesionales(inscripcion.id)
                horasvinculacion = haber_cumplido_horas_creditos_vinculacion(inscripcion.id)

                if InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallaantigua_id).exists():
                    imantigua = InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallaantigua_id)[0]
                    imantigua.status = False
                    imantigua.save()
                    print(u"Desactiva antigua inscripcion -----------------------------")

                if not InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallanueva_id).exists():
                    imnueva = InscripcionMalla(inscripcion=inscripcion, malla_id=mallanueva_id)
                    imnueva.save()
                    print(u"Crea nueva inscripcion -----------------------------")

                equivalencias = TablaEquivalenciaAsignaturas.objects.filter(status=True, asignaturamalla__malla_id=mallaantigua_id).order_by('asignaturamallasalto__nivelmalla__orden')
                cont_asig_vinculacion_aprobadas = 0
                horasfalta = 0
                fechainicioitinerario = None
                fechafinitinerario = None
                for equivalencia in equivalencias:
                    print(u"nueva - %s" % equivalencia.asignaturamallasalto)
                    recordantiguo = inscripcion.recordacademico_set.filter(status=True,asignaturamalla_id=equivalencia.asignaturamalla_id).first()

                    if recordantiguo:
                        print(u"anterior - %s" % equivalencia.asignaturamalla)
                        print(u"Record antiguo: %s" % recordantiguo)
                        recordnuevo = None
                        recordantiguo.status = False
                        recordantiguo.save(update_asignaturamalla=False)

                        if equivalencia.asignaturamallasalto_id in [11018,11021,11020,10990,10991,11005,10993,10997,11007]:
                            observaciones = recordantiguo.observaciones + " Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15"
                            homologada = True


                        else:
                            observaciones = recordantiguo.observaciones + " Migración con base a RESOLUCIÓN CGA-SO-9-2023-NO15 Y RESOLUCIÓN OCS-SO-18-2023-NO14"
                            homologada = False
                        if not RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                              asignaturamalla=equivalencia.asignaturamallasalto).exists():



                            recordnuevo = RecordAcademico(inscripcion=inscripcion,
                                                          matriculas=recordantiguo.matriculas,
                                                          asignaturamalla=equivalencia.asignaturamallasalto,
                                                          asignatura=equivalencia.asignaturamallasalto.asignatura,
                                                          asignaturaold_id=recordantiguo.asignatura.id,
                                                          nota=recordantiguo.nota,
                                                          asistencia=recordantiguo.asistencia,
                                                          sinasistencia=recordantiguo.sinasistencia,
                                                          fecha=recordantiguo.fecha,
                                                          noaplica=recordantiguo.noaplica,
                                                          aprobada=recordantiguo.aprobada,
                                                          convalidacion=recordantiguo.convalidacion,
                                                          pendiente=recordantiguo.pendiente,
                                                          creditos=equivalencia.asignaturamallasalto.creditos,
                                                          horas=equivalencia.asignaturamallasalto.horas,
                                                          valida=recordantiguo.valida,
                                                          validapromedio=recordantiguo.validapromedio,
                                                          observaciones=observaciones,
                                                          homologada=homologada,
                                                          materiaregular=recordantiguo.materiaregular,
                                                          materiacurso=None,
                                                          completonota=recordantiguo.completonota,
                                                          completoasistencia=recordantiguo.completoasistencia,
                                                          fechainicio=recordantiguo.fechainicio,
                                                          fechafin=recordantiguo.fechafin,
                                                          suficiencia=recordantiguo.suficiencia,
                                                          asignaturamallahistorico_id=recordantiguo.asignaturamalla.id,
                                                          reverso=False)
                            recordnuevo.save()
                            print(u"Crea nuevo record %s" % recordnuevo)


                        elif RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                            asignaturamalla=equivalencia.asignaturamallasalto):
                            recordnuevo = RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                         asignaturamalla=equivalencia.asignaturamallasalto)[0]
                            recordnuevo.matriculas = recordantiguo.matriculas
                            recordnuevo.asignaturamalla = equivalencia.asignaturamallasalto
                            recordnuevo.asignatura = equivalencia.asignaturamallasalto.asignatura
                            recordnuevo.asignaturaold = recordantiguo.asignaturamalla.asignatura
                            recordnuevo.nota = recordantiguo.nota
                            recordnuevo.asistencia = recordantiguo.asistencia
                            recordnuevo.sinasistencia = recordantiguo.sinasistencia
                            recordnuevo.fecha = recordantiguo.fecha
                            recordnuevo.noaplica = recordantiguo.noaplica
                            recordnuevo.aprobada = recordantiguo.aprobada
                            recordnuevo.convalidacion = recordantiguo.convalidacion
                            recordnuevo.pendiente = recordantiguo.pendiente
                            recordnuevo.creditos = equivalencia.asignaturamallasalto.creditos
                            recordnuevo.horas = equivalencia.asignaturamallasalto.horas
                            recordnuevo.valida = recordantiguo.valida
                            recordnuevo.validapromedio = recordantiguo.validapromedio
                            recordnuevo.observaciones = observaciones
                            recordnuevo.homologada = homologada
                            recordnuevo.materiaregular = recordantiguo.materiaregular
                            recordnuevo.materiacurso = None
                            recordnuevo.completonota = recordantiguo.completonota
                            recordnuevo.completoasistencia = recordantiguo.completoasistencia
                            recordnuevo.fechainicio = recordantiguo.fechainicio
                            recordnuevo.fechafin = recordantiguo.fechafin
                            recordnuevo.suficiencia = recordantiguo.suficiencia
                            recordnuevo.asignaturamallahistorico = recordantiguo.asignaturamalla
                            recordnuevo.reverso = False
                            recordnuevo.save()

                        if recordnuevo:
                            historicos = HistoricoRecordAcademico.objects.filter(status=True,
                                                                                 recordacademico=recordantiguo).update(recordacademico=recordnuevo,
                                                                                                                       creditos=recordnuevo.creditos,
                                                                                                                       horas=recordnuevo.horas,
                                                                                                                       homologada=recordnuevo.homologada)
                            respaldo = RespaldoRecordAcademico.objects.filter(status=True,recordacademicooriginal=recordantiguo)



                            if equivalencia.asignaturamallasalto_id in [11018,11021,11020,10990,10991,11005,10993,10997,11007]:
                                if not practicaspp or not horasvinculacion:
                                    if recordnuevo.aprobada:
                                        profesor = None
                                        if recordnuevo.materiaregular:
                                            profesor = recordnuevo.materiaregular.profesor_principal()
                                        elif recordnuevo.materiacurso:
                                            profesor = recordnuevo.materiaregular.profesor()
                                        if equivalencia.asignaturamallasalto_id == 11018:

                                            itinerarioprimero = ItinerariosMalla.objects.get(status=True, malla_id=mallaantigua_id, nivel_id=1)
                                            iprimeronuevo = ItinerariosMalla.objects.get(status=True, malla_id=mallanueva_id, nivel_id=1)


                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerarioprimero).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerarioprimero).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerarioprimero).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=iprimeronuevo.horas_practicas,
                                                                                                         nivelmalla=iprimeronuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=iprimeronuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerarioprimero).update(
                                                itinerario=iprimeronuevo)

                                        if equivalencia.asignaturamallasalto_id == 11021:
                                            itinerariosegundo = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=2)
                                            isegundonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=2)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerariosegundo).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerariosegundo).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerariosegundo).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=isegundonuevo.horas_practicas,
                                                                                                         nivelmalla=isegundonuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=isegundonuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerariosegundo).update(
                                                itinerario=isegundonuevo)

                                        if equivalencia.asignaturamallasalto_id == 11020:
                                            itinerariotercero = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=3)
                                            iterceronuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=3)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerariotercero).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerariotercero).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerariotercero).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=iterceronuevo.horas_practicas,
                                                                                                         nivelmalla=iterceronuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=iterceronuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerariotercero).update(
                                                itinerario=iterceronuevo)

                                        if equivalencia.asignaturamallasalto_id == 10990:
                                            itinerariocuarto = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=4)
                                            icuartonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=4)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerariocuarto).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerariocuarto).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerariocuarto).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=icuartonuevo.horas_practicas,
                                                                                                         nivelmalla=icuartonuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=icuartonuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerariocuarto).update(
                                                itinerario=icuartonuevo)

                                        if equivalencia.asignaturamallasalto_id == 10991:
                                            itinerarioquinto = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=5)
                                            iquintonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=5)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerarioquinto).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerarioquinto).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerarioquinto).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=iquintonuevo.horas_practicas,
                                                                                                         nivelmalla=iquintonuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=iquintonuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerarioquinto).update(
                                                itinerario=iquintonuevo)

                                        if equivalencia.asignaturamallasalto_id == 11005:
                                            itinerariosexto = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=6)
                                            isextonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=6)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerariosexto).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerariosexto).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerariosexto).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=isextonuevo.horas_practicas,
                                                                                                         nivelmalla=isextonuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=isextonuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerariosexto).update(
                                                itinerario=isextonuevo)

                                        if equivalencia.asignaturamallasalto_id == 10993:
                                            ####################################### VINCULACION ###################################################
                                            if not horasvinculacion:
                                                if inscripcion.numero_horas_proyectos_vinculacion() < 80:
                                                    horasfalta = 80 - inscripcion.numero_horas_proyectos_vinculacion()
                                                    vinculacion = ParticipantesMatrices(status=True,
                                                                                        matrizevidencia_id=2,
                                                                                        proyecto_id=601,
                                                                                        inscripcion=inscripcion,
                                                                                        horas=horasfalta,
                                                                                        registrohorasdesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                        registrohorashasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                        estado=1
                                                                                        )
                                                    vinculacion.save()
                                            ######################################################################################################
                                            itinerarioseptimo = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=7)
                                            iseptimonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=7)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerarioseptimo).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerarioseptimo).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerarioseptimo).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=iseptimonuevo.horas_practicas,
                                                                                                         nivelmalla=iseptimonuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=iseptimonuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerarioseptimo).update(
                                                itinerario=iseptimonuevo)

                                        if equivalencia.asignaturamallasalto_id == 10997:
                                            itinerariooctavo = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=8)
                                            ioctavonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=8)
                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerariooctavo).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerariooctavo).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerariooctavo).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=ioctavonuevo.horas_practicas,
                                                                                                         nivelmalla=ioctavonuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=ioctavonuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerariooctavo).update(
                                                itinerario=ioctavonuevo)

                                        if equivalencia.asignaturamallasalto_id == 11007:
                                            ####################################### VINCULACION ###################################################
                                            if not horasvinculacion:
                                                if inscripcion.numero_horas_proyectos_vinculacion() < 160:
                                                    horasfalta = 160 - inscripcion.numero_horas_proyectos_vinculacion()
                                                    vinculacion = ParticipantesMatrices(status=True,
                                                                                        matrizevidencia_id=2,
                                                                                        proyecto_id=601,
                                                                                        inscripcion=inscripcion,
                                                                                        horas=horasfalta,
                                                                                        registrohorasdesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                        registrohorashasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                        estado=1
                                                                                        )
                                                    vinculacion.save()
                                            ######################################################################################################
                                            itinerarionoveno = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=9)
                                            inovenonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=9)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerarionoveno).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerarionoveno).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerarionoveno).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=inovenonuevo.horas_practicas,
                                                                                                         nivelmalla=inovenonuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=inovenonuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerarionoveno).update(
                                                itinerario=inovenonuevo)


                            if not respaldo.exists():
                                respaldorecord = RespaldoRecordAcademico(
                                    recordacademicooriginal=recordantiguo,
                                    recordacademiconuevo=recordnuevo
                                )
                                respaldorecord.save()
                            else:
                                respaldorecord = respaldo[0]
                                respaldorecord.recordacademiconuevo = recordnuevo
                                respaldorecord.save()
                            print(u"Record actualizado %s" % recordnuevo)

                    else:
                        hojadestino.write(fila, 3, "NO ENCONTRO RECORD ANTIGUO %s" % equivalencia.asignaturamalla)
                        fila += 1

                practicasppf = inscripcion.numero_horas_practicas_pre_profesionales()
                hojadestino.write(fila, 3, practicasppf, fuentenormal)
                horasvinculacionf = inscripcion.numero_horas_proyectos_vinculacion()
                hojadestino.write(fila, 4, horasvinculacionf, fuentenormal)

                fila += 1

                time.sleep(1)

            lin += 1

        libdestino.save(output_folder + libre_origen)
        print(output_folder + libre_origen)
        print("Proceso finalizado. . .")

    except Exception as ex:
        transaction.set_rollback(True)
        print('error: %s' % ex)
        hojadestino.write(fila, 3, str(ex))
        fila += 1

def homologacion_idiomas3():
    #verificar que todos los estudiantes tengan la misma malla
    cadena = ''
    linea, excluidos, conexito = 0, 0, 0
    try:
        libre_origen = '/homologacion_idiomas_2_3.xls'
        fuentecabecera = easyxf(
            'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
        fuentenormal = easyxf(
            'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')

        output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
        # liborigen = xlrd.open_workbook(output_folder + libre_origen)
        libdestino = xlwt.Workbook()
        hojadestino = libdestino.add_sheet('Sheet1')
        fil = 0
        columnas = [(u"CEDULA", 6000),
                    (u"APELLIDOS Y NOMBRES", 6000),
                    (u"OBSERVACIÓN", 6000),
                    (u"HORAS PRACTICAS", 6000),
                    (u"HORAS VINCULACION", 6000),
                    (u"OBSERVACIÓN", 6000)
                    ]
        for col_num in range(len(columnas)):
            hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
            hojadestino.col(col_num).width = columnas[col_num][1]
        fila = 1
        titulacion = 0

        lin = 0
        #miarchivo = openpyxl.load_workbook("CEDULA_COMUNICACION_final.xlsx")
        miarchivo = openpyxl.load_workbook("ilinea.xlsx")
        #miarchivo = openpyxl.load_workbook("prueba_2.xlsx")

        ws = miarchivo.get_sheet_by_name("tercero")
        worksheet = ws
        c = 0
        cont = 0
        periodo_id=224
        carrera_id=129
        mallaantigua_id=198
        mallanueva_id=492

        for row in worksheet.iter_rows(min_row=0):
            if lin >= 0:
                currentValues, cadena = [], ''
                for cell in row:
                    cadena += str(cell.value) + ' '
                    currentValues.append(str(cell.value))
                identificacion = currentValues[0]

                if not identificacion:
                    break

                matricula = Matricula.objects.filter(status=True, inscripcion__carrera__id=carrera_id, inscripcion__persona__cedula=identificacion).first()
                cont += 1
                matricula.pasoayuda = True
                matricula.save()
                print(u"%s - %s" % (matricula, cont))
                inscripcion = matricula.inscripcion
                hojadestino.write(fila, 0, matricula.inscripcion.persona.identificacion(), fuentenormal)
                hojadestino.write(fila, 1, matricula.inscripcion.persona.nombre_completo(), fuentenormal)
                hojadestino.write(fila, 2, matricula.nivel.periodo.nombre, fuentenormal)

                practicaspp = haber_cumplido_horas_creditos_practicas_preprofesionales(inscripcion.id)
                horasvinculacion = haber_cumplido_horas_creditos_vinculacion(inscripcion.id)

                if InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallaantigua_id).exists():
                    imantigua = InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallaantigua_id)[0]
                    imantigua.status = False
                    imantigua.save()
                    print(u"Desactiva antigua inscripcion -----------------------------")

                if not InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallanueva_id).exists():
                    imnueva = InscripcionMalla(inscripcion=inscripcion, malla_id=mallanueva_id)
                    imnueva.save()
                    print(u"Crea nueva inscripcion -----------------------------")

                equivalencias = TablaEquivalenciaAsignaturas.objects.filter(status=True, asignaturamalla__malla_id=mallaantigua_id).order_by('asignaturamallasalto__nivelmalla__orden')
                cont_asig_vinculacion_aprobadas = 0
                horasfalta = 0
                fechainicioitinerario = None
                fechafinitinerario = None
                for equivalencia in equivalencias:
                    print(u"nueva - %s" % equivalencia.asignaturamallasalto)
                    recordantiguo = inscripcion.recordacademico_set.filter(status=True,asignaturamalla_id=equivalencia.asignaturamalla_id).first()

                    if recordantiguo:
                        print(u"anterior - %s" % equivalencia.asignaturamalla)
                        print(u"Record antiguo: %s" % recordantiguo)
                        recordnuevo = None
                        recordantiguo.status = False
                        recordantiguo.save(update_asignaturamalla=False)

                        if equivalencia.asignaturamallasalto_id in [11018,11021,11020,10990,10991,11005,10993,10997,11007]:
                            observaciones = recordantiguo.observaciones + " Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15"
                            homologada = True


                        else:
                            observaciones = recordantiguo.observaciones + " Migración con base a RESOLUCIÓN CGA-SO-9-2023-NO15 Y RESOLUCIÓN OCS-SO-18-2023-NO14"
                            homologada = False
                        if not RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                              asignaturamalla=equivalencia.asignaturamallasalto).exists():



                            recordnuevo = RecordAcademico(inscripcion=inscripcion,
                                                          matriculas=recordantiguo.matriculas,
                                                          asignaturamalla=equivalencia.asignaturamallasalto,
                                                          asignatura=equivalencia.asignaturamallasalto.asignatura,
                                                          asignaturaold_id=recordantiguo.asignatura.id,
                                                          nota=recordantiguo.nota,
                                                          asistencia=recordantiguo.asistencia,
                                                          sinasistencia=recordantiguo.sinasistencia,
                                                          fecha=recordantiguo.fecha,
                                                          noaplica=recordantiguo.noaplica,
                                                          aprobada=recordantiguo.aprobada,
                                                          convalidacion=recordantiguo.convalidacion,
                                                          pendiente=recordantiguo.pendiente,
                                                          creditos=equivalencia.asignaturamallasalto.creditos,
                                                          horas=equivalencia.asignaturamallasalto.horas,
                                                          valida=recordantiguo.valida,
                                                          validapromedio=recordantiguo.validapromedio,
                                                          observaciones=observaciones,
                                                          homologada=homologada,
                                                          materiaregular=recordantiguo.materiaregular,
                                                          materiacurso=None,
                                                          completonota=recordantiguo.completonota,
                                                          completoasistencia=recordantiguo.completoasistencia,
                                                          fechainicio=recordantiguo.fechainicio,
                                                          fechafin=recordantiguo.fechafin,
                                                          suficiencia=recordantiguo.suficiencia,
                                                          asignaturamallahistorico_id=recordantiguo.asignaturamalla.id,
                                                          reverso=False)
                            recordnuevo.save()
                            print(u"Crea nuevo record %s" % recordnuevo)


                        elif RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                            asignaturamalla=equivalencia.asignaturamallasalto):
                            recordnuevo = RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                         asignaturamalla=equivalencia.asignaturamallasalto)[0]
                            recordnuevo.matriculas = recordantiguo.matriculas
                            recordnuevo.asignaturamalla = equivalencia.asignaturamallasalto
                            recordnuevo.asignatura = equivalencia.asignaturamallasalto.asignatura
                            recordnuevo.asignaturaold = recordantiguo.asignaturamalla.asignatura
                            recordnuevo.nota = recordantiguo.nota
                            recordnuevo.asistencia = recordantiguo.asistencia
                            recordnuevo.sinasistencia = recordantiguo.sinasistencia
                            recordnuevo.fecha = recordantiguo.fecha
                            recordnuevo.noaplica = recordantiguo.noaplica
                            recordnuevo.aprobada = recordantiguo.aprobada
                            recordnuevo.convalidacion = recordantiguo.convalidacion
                            recordnuevo.pendiente = recordantiguo.pendiente
                            recordnuevo.creditos = equivalencia.asignaturamallasalto.creditos
                            recordnuevo.horas = equivalencia.asignaturamallasalto.horas
                            recordnuevo.valida = recordantiguo.valida
                            recordnuevo.validapromedio = recordantiguo.validapromedio
                            recordnuevo.observaciones = observaciones
                            recordnuevo.homologada = homologada
                            recordnuevo.materiaregular = recordantiguo.materiaregular
                            recordnuevo.materiacurso = None
                            recordnuevo.completonota = recordantiguo.completonota
                            recordnuevo.completoasistencia = recordantiguo.completoasistencia
                            recordnuevo.fechainicio = recordantiguo.fechainicio
                            recordnuevo.fechafin = recordantiguo.fechafin
                            recordnuevo.suficiencia = recordantiguo.suficiencia
                            recordnuevo.asignaturamallahistorico = recordantiguo.asignaturamalla
                            recordnuevo.reverso = False
                            recordnuevo.save()

                        if recordnuevo:
                            historicos = HistoricoRecordAcademico.objects.filter(status=True,
                                                                                 recordacademico=recordantiguo).update(recordacademico=recordnuevo,
                                                                                                                       creditos=recordnuevo.creditos,
                                                                                                                       horas=recordnuevo.horas,
                                                                                                                       homologada=recordnuevo.homologada)
                            respaldo = RespaldoRecordAcademico.objects.filter(status=True,recordacademicooriginal=recordantiguo)



                            if equivalencia.asignaturamallasalto_id in [11018,11021,11020,10990,10991,11005,10993,10997,11007]:
                                if not practicaspp or not horasvinculacion:
                                    if recordnuevo.aprobada:
                                        profesor = None
                                        if recordnuevo.materiaregular:
                                            profesor = recordnuevo.materiaregular.profesor_principal()
                                        elif recordnuevo.materiacurso:
                                            profesor = recordnuevo.materiaregular.profesor()
                                        if equivalencia.asignaturamallasalto_id == 11018:

                                            itinerarioprimero = ItinerariosMalla.objects.get(status=True, malla_id=mallaantigua_id, nivel_id=1)
                                            iprimeronuevo = ItinerariosMalla.objects.get(status=True, malla_id=mallanueva_id, nivel_id=1)


                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerarioprimero).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerarioprimero).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerarioprimero).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=iprimeronuevo.horas_practicas,
                                                                                                         nivelmalla=iprimeronuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=iprimeronuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerarioprimero).update(
                                                itinerario=iprimeronuevo)

                                        if equivalencia.asignaturamallasalto_id == 11021:
                                            itinerariosegundo = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=2)
                                            isegundonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=2)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerariosegundo).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerariosegundo).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerariosegundo).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=isegundonuevo.horas_practicas,
                                                                                                         nivelmalla=isegundonuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=isegundonuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerariosegundo).update(
                                                itinerario=isegundonuevo)

                                        if equivalencia.asignaturamallasalto_id == 11020:
                                            itinerariotercero = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=3)
                                            iterceronuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=3)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerariotercero).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerariotercero).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerariotercero).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=iterceronuevo.horas_practicas,
                                                                                                         nivelmalla=iterceronuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=iterceronuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerariotercero).update(
                                                itinerario=iterceronuevo)

                                        if equivalencia.asignaturamallasalto_id == 10990:
                                            itinerariocuarto = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=4)
                                            icuartonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=4)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerariocuarto).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerariocuarto).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerariocuarto).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=icuartonuevo.horas_practicas,
                                                                                                         nivelmalla=icuartonuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=icuartonuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerariocuarto).update(
                                                itinerario=icuartonuevo)

                                        if equivalencia.asignaturamallasalto_id == 10991:
                                            itinerarioquinto = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=5)
                                            iquintonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=5)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerarioquinto).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerarioquinto).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerarioquinto).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=iquintonuevo.horas_practicas,
                                                                                                         nivelmalla=iquintonuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=iquintonuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerarioquinto).update(
                                                itinerario=iquintonuevo)

                                        if equivalencia.asignaturamallasalto_id == 11005:
                                            itinerariosexto = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=6)
                                            isextonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=6)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerariosexto).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerariosexto).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerariosexto).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=isextonuevo.horas_practicas,
                                                                                                         nivelmalla=isextonuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=isextonuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerariosexto).update(
                                                itinerario=isextonuevo)

                                        if equivalencia.asignaturamallasalto_id == 10993:
                                            ####################################### VINCULACION ###################################################
                                            if not horasvinculacion:
                                                if inscripcion.numero_horas_proyectos_vinculacion() < 80:
                                                    horasfalta = 80 - inscripcion.numero_horas_proyectos_vinculacion()
                                                    vinculacion = ParticipantesMatrices(status=True,
                                                                                        matrizevidencia_id=2,
                                                                                        proyecto_id=601,
                                                                                        inscripcion=inscripcion,
                                                                                        horas=horasfalta,
                                                                                        registrohorasdesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                        registrohorashasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                        estado=1
                                                                                        )
                                                    vinculacion.save()
                                            ######################################################################################################
                                            itinerarioseptimo = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=7)
                                            iseptimonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=7)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerarioseptimo).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerarioseptimo).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerarioseptimo).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=iseptimonuevo.horas_practicas,
                                                                                                         nivelmalla=iseptimonuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=iseptimonuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerarioseptimo).update(
                                                itinerario=iseptimonuevo)

                                        if equivalencia.asignaturamallasalto_id == 10997:
                                            itinerariooctavo = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=8)
                                            ioctavonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=8)
                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerariooctavo).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerariooctavo).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerariooctavo).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=ioctavonuevo.horas_practicas,
                                                                                                         nivelmalla=ioctavonuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=ioctavonuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerariooctavo).update(
                                                itinerario=ioctavonuevo)

                                        if equivalencia.asignaturamallasalto_id == 11007:
                                            ####################################### VINCULACION ###################################################
                                            if not horasvinculacion:
                                                if inscripcion.numero_horas_proyectos_vinculacion() < 160:
                                                    horasfalta = 160 - inscripcion.numero_horas_proyectos_vinculacion()
                                                    vinculacion = ParticipantesMatrices(status=True,
                                                                                        matrizevidencia_id=2,
                                                                                        proyecto_id=601,
                                                                                        inscripcion=inscripcion,
                                                                                        horas=horasfalta,
                                                                                        registrohorasdesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                        registrohorashasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                        estado=1
                                                                                        )
                                                    vinculacion.save()
                                            ######################################################################################################
                                            itinerarionoveno = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=9)
                                            inovenonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=9)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerarionoveno).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerarionoveno).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerarionoveno).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=inovenonuevo.horas_practicas,
                                                                                                         nivelmalla=inovenonuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=inovenonuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerarionoveno).update(
                                                itinerario=inovenonuevo)


                            if not respaldo.exists():
                                respaldorecord = RespaldoRecordAcademico(
                                    recordacademicooriginal=recordantiguo,
                                    recordacademiconuevo=recordnuevo
                                )
                                respaldorecord.save()
                            else:
                                respaldorecord = respaldo[0]
                                respaldorecord.recordacademiconuevo = recordnuevo
                                respaldorecord.save()
                            print(u"Record actualizado %s" % recordnuevo)

                    else:
                        hojadestino.write(fila, 3, "NO ENCONTRO RECORD ANTIGUO %s" % equivalencia.asignaturamalla)
                        fila += 1

                practicasppf = inscripcion.numero_horas_practicas_pre_profesionales()
                hojadestino.write(fila, 3, practicasppf, fuentenormal)
                horasvinculacionf = inscripcion.numero_horas_proyectos_vinculacion()
                hojadestino.write(fila, 4, horasvinculacionf, fuentenormal)

                fila += 1

                time.sleep(1)

            lin += 1

        libdestino.save(output_folder + libre_origen)
        print(output_folder + libre_origen)
        print("Proceso finalizado. . .")

    except Exception as ex:
        transaction.set_rollback(True)
        print('error: %s' % ex)
        hojadestino.write(fila, 3, str(ex))
        fila += 1

def homologacion_idiomas4():
    #verificar que todos los estudiantes tengan la misma malla
    cadena = ''
    linea, excluidos, conexito = 0, 0, 0
    try:
        libre_origen = '/homologacion_idiomas_2_4.xls'
        fuentecabecera = easyxf(
            'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
        fuentenormal = easyxf(
            'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')

        output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
        # liborigen = xlrd.open_workbook(output_folder + libre_origen)
        libdestino = xlwt.Workbook()
        hojadestino = libdestino.add_sheet('Sheet1')
        fil = 0
        columnas = [(u"CEDULA", 6000),
                    (u"APELLIDOS Y NOMBRES", 6000),
                    (u"OBSERVACIÓN", 6000),
                    (u"HORAS PRACTICAS", 6000),
                    (u"HORAS VINCULACION", 6000),
                    (u"OBSERVACIÓN", 6000)
                    ]
        for col_num in range(len(columnas)):
            hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
            hojadestino.col(col_num).width = columnas[col_num][1]
        fila = 1
        titulacion = 0

        lin = 0
        #miarchivo = openpyxl.load_workbook("CEDULA_COMUNICACION_final.xlsx")
        miarchivo = openpyxl.load_workbook("ilinea.xlsx")
        #miarchivo = openpyxl.load_workbook("prueba_2.xlsx")

        ws = miarchivo.get_sheet_by_name("cuarto")
        worksheet = ws
        c = 0
        cont = 0
        periodo_id=224
        carrera_id=129
        mallaantigua_id=198
        mallanueva_id=492

        for row in worksheet.iter_rows(min_row=0):
            if lin >= 0:
                currentValues, cadena = [], ''
                for cell in row:
                    cadena += str(cell.value) + ' '
                    currentValues.append(str(cell.value))
                identificacion = currentValues[0]

                if not identificacion:
                    break

                matricula = Matricula.objects.filter(status=True, inscripcion__carrera__id=carrera_id, inscripcion__persona__cedula=identificacion).first()
                cont += 1
                matricula.pasoayuda = True
                matricula.save()
                print(u"%s - %s" % (matricula, cont))
                inscripcion = matricula.inscripcion
                hojadestino.write(fila, 0, matricula.inscripcion.persona.identificacion(), fuentenormal)
                hojadestino.write(fila, 1, matricula.inscripcion.persona.nombre_completo(), fuentenormal)
                hojadestino.write(fila, 2, matricula.nivel.periodo.nombre, fuentenormal)

                practicaspp = haber_cumplido_horas_creditos_practicas_preprofesionales(inscripcion.id)
                horasvinculacion = haber_cumplido_horas_creditos_vinculacion(inscripcion.id)

                if InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallaantigua_id).exists():
                    imantigua = InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallaantigua_id)[0]
                    imantigua.status = False
                    imantigua.save()
                    print(u"Desactiva antigua inscripcion -----------------------------")

                if not InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallanueva_id).exists():
                    imnueva = InscripcionMalla(inscripcion=inscripcion, malla_id=mallanueva_id)
                    imnueva.save()
                    print(u"Crea nueva inscripcion -----------------------------")

                equivalencias = TablaEquivalenciaAsignaturas.objects.filter(status=True, asignaturamalla__malla_id=mallaantigua_id).order_by('asignaturamallasalto__nivelmalla__orden')
                cont_asig_vinculacion_aprobadas = 0
                horasfalta = 0
                fechainicioitinerario = None
                fechafinitinerario = None
                for equivalencia in equivalencias:
                    print(u"nueva - %s" % equivalencia.asignaturamallasalto)
                    recordantiguo = inscripcion.recordacademico_set.filter(status=True,asignaturamalla_id=equivalencia.asignaturamalla_id).first()

                    if recordantiguo:
                        print(u"anterior - %s" % equivalencia.asignaturamalla)
                        print(u"Record antiguo: %s" % recordantiguo)
                        recordnuevo = None
                        recordantiguo.status = False
                        recordantiguo.save(update_asignaturamalla=False)

                        if equivalencia.asignaturamallasalto_id in [11018,11021,11020,10990,10991,11005,10993,10997,11007]:
                            observaciones = recordantiguo.observaciones + " Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15"
                            homologada = True


                        else:
                            observaciones = recordantiguo.observaciones + " Migración con base a RESOLUCIÓN CGA-SO-9-2023-NO15 Y RESOLUCIÓN OCS-SO-18-2023-NO14"
                            homologada = False
                        if not RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                              asignaturamalla=equivalencia.asignaturamallasalto).exists():



                            recordnuevo = RecordAcademico(inscripcion=inscripcion,
                                                          matriculas=recordantiguo.matriculas,
                                                          asignaturamalla=equivalencia.asignaturamallasalto,
                                                          asignatura=equivalencia.asignaturamallasalto.asignatura,
                                                          asignaturaold_id=recordantiguo.asignatura.id,
                                                          nota=recordantiguo.nota,
                                                          asistencia=recordantiguo.asistencia,
                                                          sinasistencia=recordantiguo.sinasistencia,
                                                          fecha=recordantiguo.fecha,
                                                          noaplica=recordantiguo.noaplica,
                                                          aprobada=recordantiguo.aprobada,
                                                          convalidacion=recordantiguo.convalidacion,
                                                          pendiente=recordantiguo.pendiente,
                                                          creditos=equivalencia.asignaturamallasalto.creditos,
                                                          horas=equivalencia.asignaturamallasalto.horas,
                                                          valida=recordantiguo.valida,
                                                          validapromedio=recordantiguo.validapromedio,
                                                          observaciones=observaciones,
                                                          homologada=homologada,
                                                          materiaregular=recordantiguo.materiaregular,
                                                          materiacurso=None,
                                                          completonota=recordantiguo.completonota,
                                                          completoasistencia=recordantiguo.completoasistencia,
                                                          fechainicio=recordantiguo.fechainicio,
                                                          fechafin=recordantiguo.fechafin,
                                                          suficiencia=recordantiguo.suficiencia,
                                                          asignaturamallahistorico_id=recordantiguo.asignaturamalla.id,
                                                          reverso=False)
                            recordnuevo.save()
                            print(u"Crea nuevo record %s" % recordnuevo)


                        elif RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                            asignaturamalla=equivalencia.asignaturamallasalto):
                            recordnuevo = RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                         asignaturamalla=equivalencia.asignaturamallasalto)[0]
                            recordnuevo.matriculas = recordantiguo.matriculas
                            recordnuevo.asignaturamalla = equivalencia.asignaturamallasalto
                            recordnuevo.asignatura = equivalencia.asignaturamallasalto.asignatura
                            recordnuevo.asignaturaold = recordantiguo.asignaturamalla.asignatura
                            recordnuevo.nota = recordantiguo.nota
                            recordnuevo.asistencia = recordantiguo.asistencia
                            recordnuevo.sinasistencia = recordantiguo.sinasistencia
                            recordnuevo.fecha = recordantiguo.fecha
                            recordnuevo.noaplica = recordantiguo.noaplica
                            recordnuevo.aprobada = recordantiguo.aprobada
                            recordnuevo.convalidacion = recordantiguo.convalidacion
                            recordnuevo.pendiente = recordantiguo.pendiente
                            recordnuevo.creditos = equivalencia.asignaturamallasalto.creditos
                            recordnuevo.horas = equivalencia.asignaturamallasalto.horas
                            recordnuevo.valida = recordantiguo.valida
                            recordnuevo.validapromedio = recordantiguo.validapromedio
                            recordnuevo.observaciones = observaciones
                            recordnuevo.homologada = homologada
                            recordnuevo.materiaregular = recordantiguo.materiaregular
                            recordnuevo.materiacurso = None
                            recordnuevo.completonota = recordantiguo.completonota
                            recordnuevo.completoasistencia = recordantiguo.completoasistencia
                            recordnuevo.fechainicio = recordantiguo.fechainicio
                            recordnuevo.fechafin = recordantiguo.fechafin
                            recordnuevo.suficiencia = recordantiguo.suficiencia
                            recordnuevo.asignaturamallahistorico = recordantiguo.asignaturamalla
                            recordnuevo.reverso = False
                            recordnuevo.save()

                        if recordnuevo:
                            historicos = HistoricoRecordAcademico.objects.filter(status=True,
                                                                                 recordacademico=recordantiguo).update(recordacademico=recordnuevo,
                                                                                                                       creditos=recordnuevo.creditos,
                                                                                                                       horas=recordnuevo.horas,
                                                                                                                       homologada=recordnuevo.homologada)
                            respaldo = RespaldoRecordAcademico.objects.filter(status=True,recordacademicooriginal=recordantiguo)



                            if equivalencia.asignaturamallasalto_id in [11018,11021,11020,10990,10991,11005,10993,10997,11007]:
                                if not practicaspp or not horasvinculacion:
                                    if recordnuevo.aprobada:
                                        profesor = None
                                        if recordnuevo.materiaregular:
                                            profesor = recordnuevo.materiaregular.profesor_principal()
                                        elif recordnuevo.materiacurso:
                                            profesor = recordnuevo.materiaregular.profesor()
                                        if equivalencia.asignaturamallasalto_id == 11018:

                                            itinerarioprimero = ItinerariosMalla.objects.get(status=True, malla_id=mallaantigua_id, nivel_id=1)
                                            iprimeronuevo = ItinerariosMalla.objects.get(status=True, malla_id=mallanueva_id, nivel_id=1)


                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerarioprimero).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerarioprimero).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerarioprimero).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=iprimeronuevo.horas_practicas,
                                                                                                         nivelmalla=iprimeronuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=iprimeronuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerarioprimero).update(
                                                itinerario=iprimeronuevo)

                                        if equivalencia.asignaturamallasalto_id == 11021:
                                            itinerariosegundo = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=2)
                                            isegundonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=2)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerariosegundo).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerariosegundo).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerariosegundo).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=isegundonuevo.horas_practicas,
                                                                                                         nivelmalla=isegundonuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=isegundonuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerariosegundo).update(
                                                itinerario=isegundonuevo)

                                        if equivalencia.asignaturamallasalto_id == 11020:
                                            itinerariotercero = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=3)
                                            iterceronuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=3)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerariotercero).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerariotercero).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerariotercero).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=iterceronuevo.horas_practicas,
                                                                                                         nivelmalla=iterceronuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=iterceronuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerariotercero).update(
                                                itinerario=iterceronuevo)

                                        if equivalencia.asignaturamallasalto_id == 10990:
                                            itinerariocuarto = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=4)
                                            icuartonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=4)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerariocuarto).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerariocuarto).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerariocuarto).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=icuartonuevo.horas_practicas,
                                                                                                         nivelmalla=icuartonuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=icuartonuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerariocuarto).update(
                                                itinerario=icuartonuevo)

                                        if equivalencia.asignaturamallasalto_id == 10991:
                                            itinerarioquinto = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=5)
                                            iquintonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=5)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerarioquinto).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerarioquinto).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerarioquinto).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=iquintonuevo.horas_practicas,
                                                                                                         nivelmalla=iquintonuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=iquintonuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerarioquinto).update(
                                                itinerario=iquintonuevo)

                                        if equivalencia.asignaturamallasalto_id == 11005:
                                            itinerariosexto = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=6)
                                            isextonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=6)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerariosexto).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerariosexto).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerariosexto).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=isextonuevo.horas_practicas,
                                                                                                         nivelmalla=isextonuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=isextonuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerariosexto).update(
                                                itinerario=isextonuevo)

                                        if equivalencia.asignaturamallasalto_id == 10993:
                                            ####################################### VINCULACION ###################################################
                                            if not horasvinculacion:
                                                if inscripcion.numero_horas_proyectos_vinculacion() < 80:
                                                    horasfalta = 80 - inscripcion.numero_horas_proyectos_vinculacion()
                                                    vinculacion = ParticipantesMatrices(status=True,
                                                                                        matrizevidencia_id=2,
                                                                                        proyecto_id=601,
                                                                                        inscripcion=inscripcion,
                                                                                        horas=horasfalta,
                                                                                        registrohorasdesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                        registrohorashasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                        estado=1
                                                                                        )
                                                    vinculacion.save()
                                            ######################################################################################################
                                            itinerarioseptimo = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=7)
                                            iseptimonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=7)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerarioseptimo).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerarioseptimo).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerarioseptimo).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=iseptimonuevo.horas_practicas,
                                                                                                         nivelmalla=iseptimonuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=iseptimonuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerarioseptimo).update(
                                                itinerario=iseptimonuevo)

                                        if equivalencia.asignaturamallasalto_id == 10997:
                                            itinerariooctavo = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=8)
                                            ioctavonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=8)
                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerariooctavo).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerariooctavo).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerariooctavo).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=ioctavonuevo.horas_practicas,
                                                                                                         nivelmalla=ioctavonuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=ioctavonuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerariooctavo).update(
                                                itinerario=ioctavonuevo)

                                        if equivalencia.asignaturamallasalto_id == 11007:
                                            ####################################### VINCULACION ###################################################
                                            if not horasvinculacion:
                                                if inscripcion.numero_horas_proyectos_vinculacion() < 160:
                                                    horasfalta = 160 - inscripcion.numero_horas_proyectos_vinculacion()
                                                    vinculacion = ParticipantesMatrices(status=True,
                                                                                        matrizevidencia_id=2,
                                                                                        proyecto_id=601,
                                                                                        inscripcion=inscripcion,
                                                                                        horas=horasfalta,
                                                                                        registrohorasdesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                        registrohorashasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                        estado=1
                                                                                        )
                                                    vinculacion.save()
                                            ######################################################################################################
                                            itinerarionoveno = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=9)
                                            inovenonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=9)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerarionoveno).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerarionoveno).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerarionoveno).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=inovenonuevo.horas_practicas,
                                                                                                         nivelmalla=inovenonuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=inovenonuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerarionoveno).update(
                                                itinerario=inovenonuevo)


                            if not respaldo.exists():
                                respaldorecord = RespaldoRecordAcademico(
                                    recordacademicooriginal=recordantiguo,
                                    recordacademiconuevo=recordnuevo
                                )
                                respaldorecord.save()
                            else:
                                respaldorecord = respaldo[0]
                                respaldorecord.recordacademiconuevo = recordnuevo
                                respaldorecord.save()
                            print(u"Record actualizado %s" % recordnuevo)

                    else:
                        hojadestino.write(fila, 3, "NO ENCONTRO RECORD ANTIGUO %s" % equivalencia.asignaturamalla)
                        fila += 1

                practicasppf = inscripcion.numero_horas_practicas_pre_profesionales()
                hojadestino.write(fila, 3, practicasppf, fuentenormal)
                horasvinculacionf = inscripcion.numero_horas_proyectos_vinculacion()
                hojadestino.write(fila, 4, horasvinculacionf, fuentenormal)

                fila += 1

                time.sleep(1)

            lin += 1

        libdestino.save(output_folder + libre_origen)
        print(output_folder + libre_origen)
        print("Proceso finalizado. . .")

    except Exception as ex:
        transaction.set_rollback(True)
        print('error: %s' % ex)
        hojadestino.write(fila, 3, str(ex))
        fila += 1

def homologacion_idiomas5():
    #verificar que todos los estudiantes tengan la misma malla
    cadena = ''
    linea, excluidos, conexito = 0, 0, 0
    try:
        libre_origen = '/homologacion_idiomas_2_5.xls'
        fuentecabecera = easyxf(
            'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
        fuentenormal = easyxf(
            'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')

        output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
        # liborigen = xlrd.open_workbook(output_folder + libre_origen)
        libdestino = xlwt.Workbook()
        hojadestino = libdestino.add_sheet('Sheet1')
        fil = 0
        columnas = [(u"CEDULA", 6000),
                    (u"APELLIDOS Y NOMBRES", 6000),
                    (u"OBSERVACIÓN", 6000),
                    (u"HORAS PRACTICAS", 6000),
                    (u"HORAS VINCULACION", 6000),
                    (u"OBSERVACIÓN", 6000)
                    ]
        for col_num in range(len(columnas)):
            hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
            hojadestino.col(col_num).width = columnas[col_num][1]
        fila = 1
        titulacion = 0

        lin = 0
        #miarchivo = openpyxl.load_workbook("CEDULA_COMUNICACION_final.xlsx")
        miarchivo = openpyxl.load_workbook("ilinea.xlsx")
        #miarchivo = openpyxl.load_workbook("prueba_2.xlsx")

        ws = miarchivo.get_sheet_by_name("quinto")
        worksheet = ws
        c = 0
        cont = 0
        periodo_id=224
        carrera_id=129
        mallaantigua_id=198
        mallanueva_id=492

        for row in worksheet.iter_rows(min_row=0):
            if lin >= 0:
                currentValues, cadena = [], ''
                for cell in row:
                    cadena += str(cell.value) + ' '
                    currentValues.append(str(cell.value))
                identificacion = currentValues[0]

                if not identificacion:
                    break

                matricula = Matricula.objects.filter(status=True, inscripcion__carrera__id=carrera_id, inscripcion__persona__cedula=identificacion).first()
                cont += 1
                matricula.pasoayuda = True
                matricula.save()
                print(u"%s - %s" % (matricula, cont))
                inscripcion = matricula.inscripcion
                hojadestino.write(fila, 0, matricula.inscripcion.persona.identificacion(), fuentenormal)
                hojadestino.write(fila, 1, matricula.inscripcion.persona.nombre_completo(), fuentenormal)
                hojadestino.write(fila, 2, matricula.nivel.periodo.nombre, fuentenormal)

                practicaspp = haber_cumplido_horas_creditos_practicas_preprofesionales(inscripcion.id)
                horasvinculacion = haber_cumplido_horas_creditos_vinculacion(inscripcion.id)

                if InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallaantigua_id).exists():
                    imantigua = InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallaantigua_id)[0]
                    imantigua.status = False
                    imantigua.save()
                    print(u"Desactiva antigua inscripcion -----------------------------")

                if not InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallanueva_id).exists():
                    imnueva = InscripcionMalla(inscripcion=inscripcion, malla_id=mallanueva_id)
                    imnueva.save()
                    print(u"Crea nueva inscripcion -----------------------------")

                equivalencias = TablaEquivalenciaAsignaturas.objects.filter(status=True, asignaturamalla__malla_id=mallaantigua_id).order_by('asignaturamallasalto__nivelmalla__orden')
                cont_asig_vinculacion_aprobadas = 0
                horasfalta = 0
                fechainicioitinerario = None
                fechafinitinerario = None
                for equivalencia in equivalencias:
                    print(u"nueva - %s" % equivalencia.asignaturamallasalto)
                    recordantiguo = inscripcion.recordacademico_set.filter(status=True,asignaturamalla_id=equivalencia.asignaturamalla_id).first()

                    if recordantiguo:
                        print(u"anterior - %s" % equivalencia.asignaturamalla)
                        print(u"Record antiguo: %s" % recordantiguo)
                        recordnuevo = None
                        recordantiguo.status = False
                        recordantiguo.save(update_asignaturamalla=False)

                        if equivalencia.asignaturamallasalto_id in [11018,11021,11020,10990,10991,11005,10993,10997,11007]:
                            observaciones = recordantiguo.observaciones + " Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15"
                            homologada = True


                        else:
                            observaciones = recordantiguo.observaciones + " Migración con base a RESOLUCIÓN CGA-SO-9-2023-NO15 Y RESOLUCIÓN OCS-SO-18-2023-NO14"
                            homologada = False
                        if not RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                              asignaturamalla=equivalencia.asignaturamallasalto).exists():



                            recordnuevo = RecordAcademico(inscripcion=inscripcion,
                                                          matriculas=recordantiguo.matriculas,
                                                          asignaturamalla=equivalencia.asignaturamallasalto,
                                                          asignatura=equivalencia.asignaturamallasalto.asignatura,
                                                          asignaturaold_id=recordantiguo.asignatura.id,
                                                          nota=recordantiguo.nota,
                                                          asistencia=recordantiguo.asistencia,
                                                          sinasistencia=recordantiguo.sinasistencia,
                                                          fecha=recordantiguo.fecha,
                                                          noaplica=recordantiguo.noaplica,
                                                          aprobada=recordantiguo.aprobada,
                                                          convalidacion=recordantiguo.convalidacion,
                                                          pendiente=recordantiguo.pendiente,
                                                          creditos=equivalencia.asignaturamallasalto.creditos,
                                                          horas=equivalencia.asignaturamallasalto.horas,
                                                          valida=recordantiguo.valida,
                                                          validapromedio=recordantiguo.validapromedio,
                                                          observaciones=observaciones,
                                                          homologada=homologada,
                                                          materiaregular=recordantiguo.materiaregular,
                                                          materiacurso=None,
                                                          completonota=recordantiguo.completonota,
                                                          completoasistencia=recordantiguo.completoasistencia,
                                                          fechainicio=recordantiguo.fechainicio,
                                                          fechafin=recordantiguo.fechafin,
                                                          suficiencia=recordantiguo.suficiencia,
                                                          asignaturamallahistorico_id=recordantiguo.asignaturamalla.id,
                                                          reverso=False)
                            recordnuevo.save()
                            print(u"Crea nuevo record %s" % recordnuevo)


                        elif RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                            asignaturamalla=equivalencia.asignaturamallasalto):
                            recordnuevo = RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                         asignaturamalla=equivalencia.asignaturamallasalto)[0]
                            recordnuevo.matriculas = recordantiguo.matriculas
                            recordnuevo.asignaturamalla = equivalencia.asignaturamallasalto
                            recordnuevo.asignatura = equivalencia.asignaturamallasalto.asignatura
                            recordnuevo.asignaturaold = recordantiguo.asignaturamalla.asignatura
                            recordnuevo.nota = recordantiguo.nota
                            recordnuevo.asistencia = recordantiguo.asistencia
                            recordnuevo.sinasistencia = recordantiguo.sinasistencia
                            recordnuevo.fecha = recordantiguo.fecha
                            recordnuevo.noaplica = recordantiguo.noaplica
                            recordnuevo.aprobada = recordantiguo.aprobada
                            recordnuevo.convalidacion = recordantiguo.convalidacion
                            recordnuevo.pendiente = recordantiguo.pendiente
                            recordnuevo.creditos = equivalencia.asignaturamallasalto.creditos
                            recordnuevo.horas = equivalencia.asignaturamallasalto.horas
                            recordnuevo.valida = recordantiguo.valida
                            recordnuevo.validapromedio = recordantiguo.validapromedio
                            recordnuevo.observaciones = observaciones
                            recordnuevo.homologada = homologada
                            recordnuevo.materiaregular = recordantiguo.materiaregular
                            recordnuevo.materiacurso = None
                            recordnuevo.completonota = recordantiguo.completonota
                            recordnuevo.completoasistencia = recordantiguo.completoasistencia
                            recordnuevo.fechainicio = recordantiguo.fechainicio
                            recordnuevo.fechafin = recordantiguo.fechafin
                            recordnuevo.suficiencia = recordantiguo.suficiencia
                            recordnuevo.asignaturamallahistorico = recordantiguo.asignaturamalla
                            recordnuevo.reverso = False
                            recordnuevo.save()

                        if recordnuevo:
                            historicos = HistoricoRecordAcademico.objects.filter(status=True,
                                                                                 recordacademico=recordantiguo).update(recordacademico=recordnuevo,
                                                                                                                       creditos=recordnuevo.creditos,
                                                                                                                       horas=recordnuevo.horas,
                                                                                                                       homologada=recordnuevo.homologada)
                            respaldo = RespaldoRecordAcademico.objects.filter(status=True,recordacademicooriginal=recordantiguo)



                            if equivalencia.asignaturamallasalto_id in [11018,11021,11020,10990,10991,11005,10993,10997,11007]:
                                if not practicaspp or not horasvinculacion:
                                    if recordnuevo.aprobada:
                                        profesor = None
                                        if recordnuevo.materiaregular:
                                            profesor = recordnuevo.materiaregular.profesor_principal()
                                        elif recordnuevo.materiacurso:
                                            profesor = recordnuevo.materiaregular.profesor()
                                        if equivalencia.asignaturamallasalto_id == 11018:

                                            itinerarioprimero = ItinerariosMalla.objects.get(status=True, malla_id=mallaantigua_id, nivel_id=1)
                                            iprimeronuevo = ItinerariosMalla.objects.get(status=True, malla_id=mallanueva_id, nivel_id=1)


                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerarioprimero).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerarioprimero).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerarioprimero).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=iprimeronuevo.horas_practicas,
                                                                                                         nivelmalla=iprimeronuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=iprimeronuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerarioprimero).update(
                                                itinerario=iprimeronuevo)

                                        if equivalencia.asignaturamallasalto_id == 11021:
                                            itinerariosegundo = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=2)
                                            isegundonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=2)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerariosegundo).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerariosegundo).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerariosegundo).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=isegundonuevo.horas_practicas,
                                                                                                         nivelmalla=isegundonuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=isegundonuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerariosegundo).update(
                                                itinerario=isegundonuevo)

                                        if equivalencia.asignaturamallasalto_id == 11020:
                                            itinerariotercero = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=3)
                                            iterceronuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=3)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerariotercero).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerariotercero).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerariotercero).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=iterceronuevo.horas_practicas,
                                                                                                         nivelmalla=iterceronuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=iterceronuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerariotercero).update(
                                                itinerario=iterceronuevo)

                                        if equivalencia.asignaturamallasalto_id == 10990:
                                            itinerariocuarto = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=4)
                                            icuartonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=4)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerariocuarto).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerariocuarto).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerariocuarto).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=icuartonuevo.horas_practicas,
                                                                                                         nivelmalla=icuartonuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=icuartonuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerariocuarto).update(
                                                itinerario=icuartonuevo)

                                        if equivalencia.asignaturamallasalto_id == 10991:
                                            itinerarioquinto = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=5)
                                            iquintonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=5)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerarioquinto).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerarioquinto).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerarioquinto).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=iquintonuevo.horas_practicas,
                                                                                                         nivelmalla=iquintonuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=iquintonuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerarioquinto).update(
                                                itinerario=iquintonuevo)

                                        if equivalencia.asignaturamallasalto_id == 11005:
                                            itinerariosexto = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=6)
                                            isextonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=6)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerariosexto).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerariosexto).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerariosexto).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=isextonuevo.horas_practicas,
                                                                                                         nivelmalla=isextonuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=isextonuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerariosexto).update(
                                                itinerario=isextonuevo)

                                        if equivalencia.asignaturamallasalto_id == 10993:
                                            ####################################### VINCULACION ###################################################
                                            if not horasvinculacion:
                                                if inscripcion.numero_horas_proyectos_vinculacion() < 80:
                                                    horasfalta = 80 - inscripcion.numero_horas_proyectos_vinculacion()
                                                    vinculacion = ParticipantesMatrices(status=True,
                                                                                        matrizevidencia_id=2,
                                                                                        proyecto_id=601,
                                                                                        inscripcion=inscripcion,
                                                                                        horas=horasfalta,
                                                                                        registrohorasdesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                        registrohorashasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                        estado=1
                                                                                        )
                                                    vinculacion.save()
                                            ######################################################################################################
                                            itinerarioseptimo = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=7)
                                            iseptimonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=7)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerarioseptimo).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerarioseptimo).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerarioseptimo).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=iseptimonuevo.horas_practicas,
                                                                                                         nivelmalla=iseptimonuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=iseptimonuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerarioseptimo).update(
                                                itinerario=iseptimonuevo)

                                        if equivalencia.asignaturamallasalto_id == 10997:
                                            itinerariooctavo = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=8)
                                            ioctavonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=8)
                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerariooctavo).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerariooctavo).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerariooctavo).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=ioctavonuevo.horas_practicas,
                                                                                                         nivelmalla=ioctavonuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=ioctavonuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerariooctavo).update(
                                                itinerario=ioctavonuevo)

                                        if equivalencia.asignaturamallasalto_id == 11007:
                                            ####################################### VINCULACION ###################################################
                                            if not horasvinculacion:
                                                if inscripcion.numero_horas_proyectos_vinculacion() < 160:
                                                    horasfalta = 160 - inscripcion.numero_horas_proyectos_vinculacion()
                                                    vinculacion = ParticipantesMatrices(status=True,
                                                                                        matrizevidencia_id=2,
                                                                                        proyecto_id=601,
                                                                                        inscripcion=inscripcion,
                                                                                        horas=horasfalta,
                                                                                        registrohorasdesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                        registrohorashasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                        estado=1
                                                                                        )
                                                    vinculacion.save()
                                            ######################################################################################################
                                            itinerarionoveno = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=9)
                                            inovenonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=9)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerarionoveno).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerarionoveno).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerarionoveno).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=inovenonuevo.horas_practicas,
                                                                                                         nivelmalla=inovenonuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=inovenonuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerarionoveno).update(
                                                itinerario=inovenonuevo)


                            if not respaldo.exists():
                                respaldorecord = RespaldoRecordAcademico(
                                    recordacademicooriginal=recordantiguo,
                                    recordacademiconuevo=recordnuevo
                                )
                                respaldorecord.save()
                            else:
                                respaldorecord = respaldo[0]
                                respaldorecord.recordacademiconuevo = recordnuevo
                                respaldorecord.save()
                            print(u"Record actualizado %s" % recordnuevo)

                    else:
                        hojadestino.write(fila, 3, "NO ENCONTRO RECORD ANTIGUO %s" % equivalencia.asignaturamalla)
                        fila += 1

                practicasppf = inscripcion.numero_horas_practicas_pre_profesionales()
                hojadestino.write(fila, 3, practicasppf, fuentenormal)
                horasvinculacionf = inscripcion.numero_horas_proyectos_vinculacion()
                hojadestino.write(fila, 4, horasvinculacionf, fuentenormal)

                fila += 1

                time.sleep(1)

            lin += 1

        libdestino.save(output_folder + libre_origen)
        print(output_folder + libre_origen)
        print("Proceso finalizado. . .")

    except Exception as ex:
        transaction.set_rollback(True)
        print('error: %s' % ex)
        hojadestino.write(fila, 3, str(ex))
        fila += 1

def homologacion_idiomas6():
    #verificar que todos los estudiantes tengan la misma malla
    cadena = ''
    linea, excluidos, conexito = 0, 0, 0
    try:
        libre_origen = '/homologacion_idiomas_2_6.xls'
        fuentecabecera = easyxf(
            'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
        fuentenormal = easyxf(
            'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')

        output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
        # liborigen = xlrd.open_workbook(output_folder + libre_origen)
        libdestino = xlwt.Workbook()
        hojadestino = libdestino.add_sheet('Sheet1')
        fil = 0
        columnas = [(u"CEDULA", 6000),
                    (u"APELLIDOS Y NOMBRES", 6000),
                    (u"OBSERVACIÓN", 6000),
                    (u"HORAS PRACTICAS", 6000),
                    (u"HORAS VINCULACION", 6000),
                    (u"OBSERVACIÓN", 6000)
                    ]
        for col_num in range(len(columnas)):
            hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
            hojadestino.col(col_num).width = columnas[col_num][1]
        fila = 1
        titulacion = 0

        lin = 0
        #miarchivo = openpyxl.load_workbook("CEDULA_COMUNICACION_final.xlsx")
        miarchivo = openpyxl.load_workbook("ilinea.xlsx")
        #miarchivo = openpyxl.load_workbook("prueba_2.xlsx")

        ws = miarchivo.get_sheet_by_name("sexto")
        worksheet = ws
        c = 0
        cont = 0
        periodo_id=224
        carrera_id=129
        mallaantigua_id=198
        mallanueva_id=492

        for row in worksheet.iter_rows(min_row=0):
            if lin >= 0:
                currentValues, cadena = [], ''
                for cell in row:
                    cadena += str(cell.value) + ' '
                    currentValues.append(str(cell.value))
                identificacion = currentValues[0]

                if not identificacion:
                    break

                matricula = Matricula.objects.filter(status=True, inscripcion__carrera__id=carrera_id, inscripcion__persona__cedula=identificacion).first()
                cont += 1
                matricula.pasoayuda = True
                matricula.save()
                print(u"%s - %s" % (matricula, cont))
                inscripcion = matricula.inscripcion
                hojadestino.write(fila, 0, matricula.inscripcion.persona.identificacion(), fuentenormal)
                hojadestino.write(fila, 1, matricula.inscripcion.persona.nombre_completo(), fuentenormal)
                hojadestino.write(fila, 2, matricula.nivel.periodo.nombre, fuentenormal)

                practicaspp = haber_cumplido_horas_creditos_practicas_preprofesionales(inscripcion.id)
                horasvinculacion = haber_cumplido_horas_creditos_vinculacion(inscripcion.id)

                if InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallaantigua_id).exists():
                    imantigua = InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallaantigua_id)[0]
                    imantigua.status = False
                    imantigua.save()
                    print(u"Desactiva antigua inscripcion -----------------------------")

                if not InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallanueva_id).exists():
                    imnueva = InscripcionMalla(inscripcion=inscripcion, malla_id=mallanueva_id)
                    imnueva.save()
                    print(u"Crea nueva inscripcion -----------------------------")

                equivalencias = TablaEquivalenciaAsignaturas.objects.filter(status=True, asignaturamalla__malla_id=mallaantigua_id).order_by('asignaturamallasalto__nivelmalla__orden')
                cont_asig_vinculacion_aprobadas = 0
                horasfalta = 0
                fechainicioitinerario = None
                fechafinitinerario = None
                for equivalencia in equivalencias:
                    print(u"nueva - %s" % equivalencia.asignaturamallasalto)
                    recordantiguo = inscripcion.recordacademico_set.filter(status=True,asignaturamalla_id=equivalencia.asignaturamalla_id).first()

                    if recordantiguo:
                        print(u"anterior - %s" % equivalencia.asignaturamalla)
                        print(u"Record antiguo: %s" % recordantiguo)
                        recordnuevo = None
                        recordantiguo.status = False
                        recordantiguo.save(update_asignaturamalla=False)

                        if equivalencia.asignaturamallasalto_id in [11018,11021,11020,10990,10991,11005,10993,10997,11007]:
                            observaciones = recordantiguo.observaciones + " Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15"
                            homologada = True


                        else:
                            observaciones = recordantiguo.observaciones + " Migración con base a RESOLUCIÓN CGA-SO-9-2023-NO15 Y RESOLUCIÓN OCS-SO-18-2023-NO14"
                            homologada = False
                        if not RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                              asignaturamalla=equivalencia.asignaturamallasalto).exists():



                            recordnuevo = RecordAcademico(inscripcion=inscripcion,
                                                          matriculas=recordantiguo.matriculas,
                                                          asignaturamalla=equivalencia.asignaturamallasalto,
                                                          asignatura=equivalencia.asignaturamallasalto.asignatura,
                                                          asignaturaold_id=recordantiguo.asignatura.id,
                                                          nota=recordantiguo.nota,
                                                          asistencia=recordantiguo.asistencia,
                                                          sinasistencia=recordantiguo.sinasistencia,
                                                          fecha=recordantiguo.fecha,
                                                          noaplica=recordantiguo.noaplica,
                                                          aprobada=recordantiguo.aprobada,
                                                          convalidacion=recordantiguo.convalidacion,
                                                          pendiente=recordantiguo.pendiente,
                                                          creditos=equivalencia.asignaturamallasalto.creditos,
                                                          horas=equivalencia.asignaturamallasalto.horas,
                                                          valida=recordantiguo.valida,
                                                          validapromedio=recordantiguo.validapromedio,
                                                          observaciones=observaciones,
                                                          homologada=homologada,
                                                          materiaregular=recordantiguo.materiaregular,
                                                          materiacurso=None,
                                                          completonota=recordantiguo.completonota,
                                                          completoasistencia=recordantiguo.completoasistencia,
                                                          fechainicio=recordantiguo.fechainicio,
                                                          fechafin=recordantiguo.fechafin,
                                                          suficiencia=recordantiguo.suficiencia,
                                                          asignaturamallahistorico_id=recordantiguo.asignaturamalla.id,
                                                          reverso=False)
                            recordnuevo.save()
                            print(u"Crea nuevo record %s" % recordnuevo)


                        elif RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                            asignaturamalla=equivalencia.asignaturamallasalto):
                            recordnuevo = RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                         asignaturamalla=equivalencia.asignaturamallasalto)[0]
                            recordnuevo.matriculas = recordantiguo.matriculas
                            recordnuevo.asignaturamalla = equivalencia.asignaturamallasalto
                            recordnuevo.asignatura = equivalencia.asignaturamallasalto.asignatura
                            recordnuevo.asignaturaold = recordantiguo.asignaturamalla.asignatura
                            recordnuevo.nota = recordantiguo.nota
                            recordnuevo.asistencia = recordantiguo.asistencia
                            recordnuevo.sinasistencia = recordantiguo.sinasistencia
                            recordnuevo.fecha = recordantiguo.fecha
                            recordnuevo.noaplica = recordantiguo.noaplica
                            recordnuevo.aprobada = recordantiguo.aprobada
                            recordnuevo.convalidacion = recordantiguo.convalidacion
                            recordnuevo.pendiente = recordantiguo.pendiente
                            recordnuevo.creditos = equivalencia.asignaturamallasalto.creditos
                            recordnuevo.horas = equivalencia.asignaturamallasalto.horas
                            recordnuevo.valida = recordantiguo.valida
                            recordnuevo.validapromedio = recordantiguo.validapromedio
                            recordnuevo.observaciones = observaciones
                            recordnuevo.homologada = homologada
                            recordnuevo.materiaregular = recordantiguo.materiaregular
                            recordnuevo.materiacurso = None
                            recordnuevo.completonota = recordantiguo.completonota
                            recordnuevo.completoasistencia = recordantiguo.completoasistencia
                            recordnuevo.fechainicio = recordantiguo.fechainicio
                            recordnuevo.fechafin = recordantiguo.fechafin
                            recordnuevo.suficiencia = recordantiguo.suficiencia
                            recordnuevo.asignaturamallahistorico = recordantiguo.asignaturamalla
                            recordnuevo.reverso = False
                            recordnuevo.save()

                        if recordnuevo:
                            historicos = HistoricoRecordAcademico.objects.filter(status=True,
                                                                                 recordacademico=recordantiguo).update(recordacademico=recordnuevo,
                                                                                                                       creditos=recordnuevo.creditos,
                                                                                                                       horas=recordnuevo.horas,
                                                                                                                       homologada=recordnuevo.homologada)
                            respaldo = RespaldoRecordAcademico.objects.filter(status=True,recordacademicooriginal=recordantiguo)



                            if equivalencia.asignaturamallasalto_id in [11018,11021,11020,10990,10991,11005,10993,10997,11007]:
                                if not practicaspp or not horasvinculacion:
                                    if recordnuevo.aprobada:
                                        profesor = None
                                        if recordnuevo.materiaregular:
                                            profesor = recordnuevo.materiaregular.profesor_principal()
                                        elif recordnuevo.materiacurso:
                                            profesor = recordnuevo.materiaregular.profesor()
                                        if equivalencia.asignaturamallasalto_id == 11018:

                                            itinerarioprimero = ItinerariosMalla.objects.get(status=True, malla_id=mallaantigua_id, nivel_id=1)
                                            iprimeronuevo = ItinerariosMalla.objects.get(status=True, malla_id=mallanueva_id, nivel_id=1)


                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerarioprimero).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerarioprimero).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerarioprimero).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=iprimeronuevo.horas_practicas,
                                                                                                         nivelmalla=iprimeronuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=iprimeronuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerarioprimero).update(
                                                itinerario=iprimeronuevo)

                                        if equivalencia.asignaturamallasalto_id == 11021:
                                            itinerariosegundo = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=2)
                                            isegundonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=2)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerariosegundo).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerariosegundo).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerariosegundo).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=isegundonuevo.horas_practicas,
                                                                                                         nivelmalla=isegundonuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=isegundonuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerariosegundo).update(
                                                itinerario=isegundonuevo)

                                        if equivalencia.asignaturamallasalto_id == 11020:
                                            itinerariotercero = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=3)
                                            iterceronuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=3)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerariotercero).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerariotercero).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerariotercero).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=iterceronuevo.horas_practicas,
                                                                                                         nivelmalla=iterceronuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=iterceronuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerariotercero).update(
                                                itinerario=iterceronuevo)

                                        if equivalencia.asignaturamallasalto_id == 10990:
                                            itinerariocuarto = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=4)
                                            icuartonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=4)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerariocuarto).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerariocuarto).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerariocuarto).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=icuartonuevo.horas_practicas,
                                                                                                         nivelmalla=icuartonuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=icuartonuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerariocuarto).update(
                                                itinerario=icuartonuevo)

                                        if equivalencia.asignaturamallasalto_id == 10991:
                                            itinerarioquinto = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=5)
                                            iquintonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=5)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerarioquinto).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerarioquinto).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerarioquinto).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=iquintonuevo.horas_practicas,
                                                                                                         nivelmalla=iquintonuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=iquintonuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerarioquinto).update(
                                                itinerario=iquintonuevo)

                                        if equivalencia.asignaturamallasalto_id == 11005:
                                            itinerariosexto = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=6)
                                            isextonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=6)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerariosexto).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerariosexto).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerariosexto).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=isextonuevo.horas_practicas,
                                                                                                         nivelmalla=isextonuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=isextonuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerariosexto).update(
                                                itinerario=isextonuevo)

                                        if equivalencia.asignaturamallasalto_id == 10993:
                                            ####################################### VINCULACION ###################################################
                                            if not horasvinculacion:
                                                if inscripcion.numero_horas_proyectos_vinculacion() < 80:
                                                    horasfalta = 80 - inscripcion.numero_horas_proyectos_vinculacion()
                                                    vinculacion = ParticipantesMatrices(status=True,
                                                                                        matrizevidencia_id=2,
                                                                                        proyecto_id=601,
                                                                                        inscripcion=inscripcion,
                                                                                        horas=horasfalta,
                                                                                        registrohorasdesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                        registrohorashasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                        estado=1
                                                                                        )
                                                    vinculacion.save()
                                            ######################################################################################################
                                            itinerarioseptimo = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=7)
                                            iseptimonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=7)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerarioseptimo).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerarioseptimo).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerarioseptimo).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=iseptimonuevo.horas_practicas,
                                                                                                         nivelmalla=iseptimonuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=iseptimonuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerarioseptimo).update(
                                                itinerario=iseptimonuevo)

                                        if equivalencia.asignaturamallasalto_id == 10997:
                                            itinerariooctavo = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=8)
                                            ioctavonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=8)
                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerariooctavo).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerariooctavo).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerariooctavo).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=ioctavonuevo.horas_practicas,
                                                                                                         nivelmalla=ioctavonuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=ioctavonuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerariooctavo).update(
                                                itinerario=ioctavonuevo)

                                        if equivalencia.asignaturamallasalto_id == 11007:
                                            ####################################### VINCULACION ###################################################
                                            if not horasvinculacion:
                                                if inscripcion.numero_horas_proyectos_vinculacion() < 160:
                                                    horasfalta = 160 - inscripcion.numero_horas_proyectos_vinculacion()
                                                    vinculacion = ParticipantesMatrices(status=True,
                                                                                        matrizevidencia_id=2,
                                                                                        proyecto_id=601,
                                                                                        inscripcion=inscripcion,
                                                                                        horas=horasfalta,
                                                                                        registrohorasdesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                        registrohorashasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                        estado=1
                                                                                        )
                                                    vinculacion.save()
                                            ######################################################################################################
                                            itinerarionoveno = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=9)
                                            inovenonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=9)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerarionoveno).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerarionoveno).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerarionoveno).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=inovenonuevo.horas_practicas,
                                                                                                         nivelmalla=inovenonuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=inovenonuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerarionoveno).update(
                                                itinerario=inovenonuevo)


                            if not respaldo.exists():
                                respaldorecord = RespaldoRecordAcademico(
                                    recordacademicooriginal=recordantiguo,
                                    recordacademiconuevo=recordnuevo
                                )
                                respaldorecord.save()
                            else:
                                respaldorecord = respaldo[0]
                                respaldorecord.recordacademiconuevo = recordnuevo
                                respaldorecord.save()
                            print(u"Record actualizado %s" % recordnuevo)

                    else:
                        hojadestino.write(fila, 3, "NO ENCONTRO RECORD ANTIGUO %s" % equivalencia.asignaturamalla)
                        fila += 1

                practicasppf = inscripcion.numero_horas_practicas_pre_profesionales()
                hojadestino.write(fila, 3, practicasppf, fuentenormal)
                horasvinculacionf = inscripcion.numero_horas_proyectos_vinculacion()
                hojadestino.write(fila, 4, horasvinculacionf, fuentenormal)

                fila += 1

                time.sleep(1)

            lin += 1

        libdestino.save(output_folder + libre_origen)
        print(output_folder + libre_origen)
        print("Proceso finalizado. . .")

    except Exception as ex:
        transaction.set_rollback(True)
        print('error: %s' % ex)
        hojadestino.write(fila, 3, str(ex))
        fila += 1

def homologacion_idiomas7():
    #verificar que todos los estudiantes tengan la misma malla
    cadena = ''
    linea, excluidos, conexito = 0, 0, 0
    try:
        libre_origen = '/homologacion_idiomas_2_7.xls'
        fuentecabecera = easyxf(
            'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
        fuentenormal = easyxf(
            'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')

        output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
        # liborigen = xlrd.open_workbook(output_folder + libre_origen)
        libdestino = xlwt.Workbook()
        hojadestino = libdestino.add_sheet('Sheet1')
        fil = 0
        columnas = [(u"CEDULA", 6000),
                    (u"APELLIDOS Y NOMBRES", 6000),
                    (u"OBSERVACIÓN", 6000),
                    (u"HORAS PRACTICAS", 6000),
                    (u"HORAS VINCULACION", 6000),
                    (u"OBSERVACIÓN", 6000)
                    ]
        for col_num in range(len(columnas)):
            hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
            hojadestino.col(col_num).width = columnas[col_num][1]
        fila = 1
        titulacion = 0

        lin = 0
        #miarchivo = openpyxl.load_workbook("CEDULA_COMUNICACION_final.xlsx")
        miarchivo = openpyxl.load_workbook("ilinea.xlsx")
        #miarchivo = openpyxl.load_workbook("prueba_2.xlsx")

        ws = miarchivo.get_sheet_by_name("septimo")
        worksheet = ws
        c = 0
        cont = 0
        periodo_id=224
        carrera_id=129
        mallaantigua_id=198
        mallanueva_id=492

        for row in worksheet.iter_rows(min_row=0):
            if lin >= 0:
                currentValues, cadena = [], ''
                for cell in row:
                    cadena += str(cell.value) + ' '
                    currentValues.append(str(cell.value))
                identificacion = currentValues[0]

                if not identificacion:
                    break

                matricula = Matricula.objects.filter(status=True, inscripcion__carrera__id=carrera_id, inscripcion__persona__cedula=identificacion).first()
                cont += 1
                matricula.pasoayuda = True
                matricula.save()
                print(u"%s - %s" % (matricula, cont))
                inscripcion = matricula.inscripcion
                hojadestino.write(fila, 0, matricula.inscripcion.persona.identificacion(), fuentenormal)
                hojadestino.write(fila, 1, matricula.inscripcion.persona.nombre_completo(), fuentenormal)
                hojadestino.write(fila, 2, matricula.nivel.periodo.nombre, fuentenormal)

                practicaspp = haber_cumplido_horas_creditos_practicas_preprofesionales(inscripcion.id)
                horasvinculacion = haber_cumplido_horas_creditos_vinculacion(inscripcion.id)

                if InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallaantigua_id).exists():
                    imantigua = InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallaantigua_id)[0]
                    imantigua.status = False
                    imantigua.save()
                    print(u"Desactiva antigua inscripcion -----------------------------")

                if not InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallanueva_id).exists():
                    imnueva = InscripcionMalla(inscripcion=inscripcion, malla_id=mallanueva_id)
                    imnueva.save()
                    print(u"Crea nueva inscripcion -----------------------------")

                equivalencias = TablaEquivalenciaAsignaturas.objects.filter(status=True, asignaturamalla__malla_id=mallaantigua_id).order_by('asignaturamallasalto__nivelmalla__orden')
                cont_asig_vinculacion_aprobadas = 0
                horasfalta = 0
                fechainicioitinerario = None
                fechafinitinerario = None
                for equivalencia in equivalencias:
                    print(u"nueva - %s" % equivalencia.asignaturamallasalto)
                    recordantiguo = inscripcion.recordacademico_set.filter(status=True,asignaturamalla_id=equivalencia.asignaturamalla_id).first()

                    if recordantiguo:
                        print(u"anterior - %s" % equivalencia.asignaturamalla)
                        print(u"Record antiguo: %s" % recordantiguo)
                        recordnuevo = None
                        recordantiguo.status = False
                        recordantiguo.save(update_asignaturamalla=False)

                        if equivalencia.asignaturamallasalto_id in [11018,11021,11020,10990,10991,11005,10993,10997,11007]:
                            observaciones = recordantiguo.observaciones + " Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15"
                            homologada = True


                        else:
                            observaciones = recordantiguo.observaciones + " Migración con base a RESOLUCIÓN CGA-SO-9-2023-NO15 Y RESOLUCIÓN OCS-SO-18-2023-NO14"
                            homologada = False
                        if not RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                              asignaturamalla=equivalencia.asignaturamallasalto).exists():



                            recordnuevo = RecordAcademico(inscripcion=inscripcion,
                                                          matriculas=recordantiguo.matriculas,
                                                          asignaturamalla=equivalencia.asignaturamallasalto,
                                                          asignatura=equivalencia.asignaturamallasalto.asignatura,
                                                          asignaturaold_id=recordantiguo.asignatura.id,
                                                          nota=recordantiguo.nota,
                                                          asistencia=recordantiguo.asistencia,
                                                          sinasistencia=recordantiguo.sinasistencia,
                                                          fecha=recordantiguo.fecha,
                                                          noaplica=recordantiguo.noaplica,
                                                          aprobada=recordantiguo.aprobada,
                                                          convalidacion=recordantiguo.convalidacion,
                                                          pendiente=recordantiguo.pendiente,
                                                          creditos=equivalencia.asignaturamallasalto.creditos,
                                                          horas=equivalencia.asignaturamallasalto.horas,
                                                          valida=recordantiguo.valida,
                                                          validapromedio=recordantiguo.validapromedio,
                                                          observaciones=observaciones,
                                                          homologada=homologada,
                                                          materiaregular=recordantiguo.materiaregular,
                                                          materiacurso=None,
                                                          completonota=recordantiguo.completonota,
                                                          completoasistencia=recordantiguo.completoasistencia,
                                                          fechainicio=recordantiguo.fechainicio,
                                                          fechafin=recordantiguo.fechafin,
                                                          suficiencia=recordantiguo.suficiencia,
                                                          asignaturamallahistorico_id=recordantiguo.asignaturamalla.id,
                                                          reverso=False)
                            recordnuevo.save()
                            print(u"Crea nuevo record %s" % recordnuevo)


                        elif RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                            asignaturamalla=equivalencia.asignaturamallasalto):
                            recordnuevo = RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                         asignaturamalla=equivalencia.asignaturamallasalto)[0]
                            recordnuevo.matriculas = recordantiguo.matriculas
                            recordnuevo.asignaturamalla = equivalencia.asignaturamallasalto
                            recordnuevo.asignatura = equivalencia.asignaturamallasalto.asignatura
                            recordnuevo.asignaturaold = recordantiguo.asignaturamalla.asignatura
                            recordnuevo.nota = recordantiguo.nota
                            recordnuevo.asistencia = recordantiguo.asistencia
                            recordnuevo.sinasistencia = recordantiguo.sinasistencia
                            recordnuevo.fecha = recordantiguo.fecha
                            recordnuevo.noaplica = recordantiguo.noaplica
                            recordnuevo.aprobada = recordantiguo.aprobada
                            recordnuevo.convalidacion = recordantiguo.convalidacion
                            recordnuevo.pendiente = recordantiguo.pendiente
                            recordnuevo.creditos = equivalencia.asignaturamallasalto.creditos
                            recordnuevo.horas = equivalencia.asignaturamallasalto.horas
                            recordnuevo.valida = recordantiguo.valida
                            recordnuevo.validapromedio = recordantiguo.validapromedio
                            recordnuevo.observaciones = observaciones
                            recordnuevo.homologada = homologada
                            recordnuevo.materiaregular = recordantiguo.materiaregular
                            recordnuevo.materiacurso = None
                            recordnuevo.completonota = recordantiguo.completonota
                            recordnuevo.completoasistencia = recordantiguo.completoasistencia
                            recordnuevo.fechainicio = recordantiguo.fechainicio
                            recordnuevo.fechafin = recordantiguo.fechafin
                            recordnuevo.suficiencia = recordantiguo.suficiencia
                            recordnuevo.asignaturamallahistorico = recordantiguo.asignaturamalla
                            recordnuevo.reverso = False
                            recordnuevo.save()

                        if recordnuevo:
                            historicos = HistoricoRecordAcademico.objects.filter(status=True,
                                                                                 recordacademico=recordantiguo).update(recordacademico=recordnuevo,
                                                                                                                       creditos=recordnuevo.creditos,
                                                                                                                       horas=recordnuevo.horas,
                                                                                                                       homologada=recordnuevo.homologada)
                            respaldo = RespaldoRecordAcademico.objects.filter(status=True,recordacademicooriginal=recordantiguo)



                            if equivalencia.asignaturamallasalto_id in [11018,11021,11020,10990,10991,11005,10993,10997,11007]:
                                if not practicaspp or not horasvinculacion:
                                    if recordnuevo.aprobada:
                                        profesor = None
                                        if recordnuevo.materiaregular:
                                            profesor = recordnuevo.materiaregular.profesor_principal()
                                        elif recordnuevo.materiacurso:
                                            profesor = recordnuevo.materiaregular.profesor()
                                        if equivalencia.asignaturamallasalto_id == 11018:

                                            itinerarioprimero = ItinerariosMalla.objects.get(status=True, malla_id=mallaantigua_id, nivel_id=1)
                                            iprimeronuevo = ItinerariosMalla.objects.get(status=True, malla_id=mallanueva_id, nivel_id=1)


                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerarioprimero).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerarioprimero).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerarioprimero).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=iprimeronuevo.horas_practicas,
                                                                                                         nivelmalla=iprimeronuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=iprimeronuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerarioprimero).update(
                                                itinerario=iprimeronuevo)

                                        if equivalencia.asignaturamallasalto_id == 11021:
                                            itinerariosegundo = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=2)
                                            isegundonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=2)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerariosegundo).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerariosegundo).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerariosegundo).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=isegundonuevo.horas_practicas,
                                                                                                         nivelmalla=isegundonuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=isegundonuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerariosegundo).update(
                                                itinerario=isegundonuevo)

                                        if equivalencia.asignaturamallasalto_id == 11020:
                                            itinerariotercero = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=3)
                                            iterceronuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=3)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerariotercero).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerariotercero).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerariotercero).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=iterceronuevo.horas_practicas,
                                                                                                         nivelmalla=iterceronuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=iterceronuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerariotercero).update(
                                                itinerario=iterceronuevo)

                                        if equivalencia.asignaturamallasalto_id == 10990:
                                            itinerariocuarto = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=4)
                                            icuartonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=4)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerariocuarto).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerariocuarto).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerariocuarto).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=icuartonuevo.horas_practicas,
                                                                                                         nivelmalla=icuartonuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=icuartonuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerariocuarto).update(
                                                itinerario=icuartonuevo)

                                        if equivalencia.asignaturamallasalto_id == 10991:
                                            itinerarioquinto = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=5)
                                            iquintonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=5)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerarioquinto).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerarioquinto).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerarioquinto).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=iquintonuevo.horas_practicas,
                                                                                                         nivelmalla=iquintonuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=iquintonuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerarioquinto).update(
                                                itinerario=iquintonuevo)

                                        if equivalencia.asignaturamallasalto_id == 11005:
                                            itinerariosexto = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=6)
                                            isextonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=6)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerariosexto).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerariosexto).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerariosexto).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=isextonuevo.horas_practicas,
                                                                                                         nivelmalla=isextonuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=isextonuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerariosexto).update(
                                                itinerario=isextonuevo)

                                        if equivalencia.asignaturamallasalto_id == 10993:
                                            ####################################### VINCULACION ###################################################
                                            if not horasvinculacion:
                                                if inscripcion.numero_horas_proyectos_vinculacion() < 80:
                                                    horasfalta = 80 - inscripcion.numero_horas_proyectos_vinculacion()
                                                    vinculacion = ParticipantesMatrices(status=True,
                                                                                        matrizevidencia_id=2,
                                                                                        proyecto_id=601,
                                                                                        inscripcion=inscripcion,
                                                                                        horas=horasfalta,
                                                                                        registrohorasdesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                        registrohorashasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                        estado=1
                                                                                        )
                                                    vinculacion.save()
                                            ######################################################################################################
                                            itinerarioseptimo = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=7)
                                            iseptimonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=7)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerarioseptimo).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerarioseptimo).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerarioseptimo).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=iseptimonuevo.horas_practicas,
                                                                                                         nivelmalla=iseptimonuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=iseptimonuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerarioseptimo).update(
                                                itinerario=iseptimonuevo)

                                        if equivalencia.asignaturamallasalto_id == 10997:
                                            itinerariooctavo = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=8)
                                            ioctavonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=8)
                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerariooctavo).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerariooctavo).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerariooctavo).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=ioctavonuevo.horas_practicas,
                                                                                                         nivelmalla=ioctavonuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=ioctavonuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerariooctavo).update(
                                                itinerario=ioctavonuevo)

                                        if equivalencia.asignaturamallasalto_id == 11007:
                                            ####################################### VINCULACION ###################################################
                                            if not horasvinculacion:
                                                if inscripcion.numero_horas_proyectos_vinculacion() < 160:
                                                    horasfalta = 160 - inscripcion.numero_horas_proyectos_vinculacion()
                                                    vinculacion = ParticipantesMatrices(status=True,
                                                                                        matrizevidencia_id=2,
                                                                                        proyecto_id=601,
                                                                                        inscripcion=inscripcion,
                                                                                        horas=horasfalta,
                                                                                        registrohorasdesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                        registrohorashasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                        estado=1
                                                                                        )
                                                    vinculacion.save()
                                            ######################################################################################################
                                            itinerarionoveno = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=9)
                                            inovenonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=9)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerarionoveno).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerarionoveno).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerarionoveno).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=inovenonuevo.horas_practicas,
                                                                                                         nivelmalla=inovenonuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=inovenonuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerarionoveno).update(
                                                itinerario=inovenonuevo)


                            if not respaldo.exists():
                                respaldorecord = RespaldoRecordAcademico(
                                    recordacademicooriginal=recordantiguo,
                                    recordacademiconuevo=recordnuevo
                                )
                                respaldorecord.save()
                            else:
                                respaldorecord = respaldo[0]
                                respaldorecord.recordacademiconuevo = recordnuevo
                                respaldorecord.save()
                            print(u"Record actualizado %s" % recordnuevo)

                    else:
                        hojadestino.write(fila, 3, "NO ENCONTRO RECORD ANTIGUO %s" % equivalencia.asignaturamalla)
                        fila += 1

                practicasppf = inscripcion.numero_horas_practicas_pre_profesionales()
                hojadestino.write(fila, 3, practicasppf, fuentenormal)
                horasvinculacionf = inscripcion.numero_horas_proyectos_vinculacion()
                hojadestino.write(fila, 4, horasvinculacionf, fuentenormal)

                fila += 1

                time.sleep(1)

            lin += 1

        libdestino.save(output_folder + libre_origen)
        print(output_folder + libre_origen)
        print("Proceso finalizado. . .")

    except Exception as ex:
        transaction.set_rollback(True)
        print('error: %s' % ex)
        hojadestino.write(fila, 3, str(ex))
        fila += 1

def homologacion_idiomas8():
    #verificar que todos los estudiantes tengan la misma malla
    cadena = ''
    linea, excluidos, conexito = 0, 0, 0
    try:
        libre_origen = '/homologacion_idiomas_2_8.xls'
        fuentecabecera = easyxf(
            'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
        fuentenormal = easyxf(
            'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')

        output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
        # liborigen = xlrd.open_workbook(output_folder + libre_origen)
        libdestino = xlwt.Workbook()
        hojadestino = libdestino.add_sheet('Sheet1')
        fil = 0
        columnas = [(u"CEDULA", 6000),
                    (u"APELLIDOS Y NOMBRES", 6000),
                    (u"OBSERVACIÓN", 6000),
                    (u"HORAS PRACTICAS", 6000),
                    (u"HORAS VINCULACION", 6000),
                    (u"OBSERVACIÓN", 6000)
                    ]
        for col_num in range(len(columnas)):
            hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
            hojadestino.col(col_num).width = columnas[col_num][1]
        fila = 1
        titulacion = 0

        lin = 0
        #miarchivo = openpyxl.load_workbook("CEDULA_COMUNICACION_final.xlsx")
        miarchivo = openpyxl.load_workbook("ilinea.xlsx")
        #miarchivo = openpyxl.load_workbook("prueba_2.xlsx")

        ws = miarchivo.get_sheet_by_name("octavo")
        worksheet = ws
        c = 0
        cont = 0
        periodo_id=224
        carrera_id=129
        mallaantigua_id=198
        mallanueva_id=492

        for row in worksheet.iter_rows(min_row=0):
            if lin >= 0:
                currentValues, cadena = [], ''
                for cell in row:
                    cadena += str(cell.value) + ' '
                    currentValues.append(str(cell.value))
                identificacion = currentValues[0]

                if not identificacion:
                    break

                matricula = Matricula.objects.filter(status=True, inscripcion__carrera__id=carrera_id, inscripcion__persona__cedula=identificacion).first()
                cont += 1
                matricula.pasoayuda = True
                matricula.save()
                print(u"%s - %s" % (matricula, cont))
                inscripcion = matricula.inscripcion
                hojadestino.write(fila, 0, matricula.inscripcion.persona.identificacion(), fuentenormal)
                hojadestino.write(fila, 1, matricula.inscripcion.persona.nombre_completo(), fuentenormal)
                hojadestino.write(fila, 2, matricula.nivel.periodo.nombre, fuentenormal)

                practicaspp = haber_cumplido_horas_creditos_practicas_preprofesionales(inscripcion.id)
                horasvinculacion = haber_cumplido_horas_creditos_vinculacion(inscripcion.id)

                if InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallaantigua_id).exists():
                    imantigua = InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallaantigua_id)[0]
                    imantigua.status = False
                    imantigua.save()
                    print(u"Desactiva antigua inscripcion -----------------------------")

                if not InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallanueva_id).exists():
                    imnueva = InscripcionMalla(inscripcion=inscripcion, malla_id=mallanueva_id)
                    imnueva.save()
                    print(u"Crea nueva inscripcion -----------------------------")

                equivalencias = TablaEquivalenciaAsignaturas.objects.filter(status=True, asignaturamalla__malla_id=mallaantigua_id).order_by('asignaturamallasalto__nivelmalla__orden')
                cont_asig_vinculacion_aprobadas = 0
                horasfalta = 0
                fechainicioitinerario = None
                fechafinitinerario = None
                for equivalencia in equivalencias:
                    print(u"nueva - %s" % equivalencia.asignaturamallasalto)
                    recordantiguo = inscripcion.recordacademico_set.filter(status=True,asignaturamalla_id=equivalencia.asignaturamalla_id).first()

                    if recordantiguo:
                        print(u"anterior - %s" % equivalencia.asignaturamalla)
                        print(u"Record antiguo: %s" % recordantiguo)
                        recordnuevo = None
                        recordantiguo.status = False
                        recordantiguo.save(update_asignaturamalla=False)

                        if equivalencia.asignaturamallasalto_id in [11018,11021,11020,10990,10991,11005,10993,10997,11007]:
                            observaciones = recordantiguo.observaciones + " Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15"
                            homologada = True


                        else:
                            observaciones = recordantiguo.observaciones + " Migración con base a RESOLUCIÓN CGA-SO-9-2023-NO15 Y RESOLUCIÓN OCS-SO-18-2023-NO14"
                            homologada = False
                        if not RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                              asignaturamalla=equivalencia.asignaturamallasalto).exists():



                            recordnuevo = RecordAcademico(inscripcion=inscripcion,
                                                          matriculas=recordantiguo.matriculas,
                                                          asignaturamalla=equivalencia.asignaturamallasalto,
                                                          asignatura=equivalencia.asignaturamallasalto.asignatura,
                                                          asignaturaold_id=recordantiguo.asignatura.id,
                                                          nota=recordantiguo.nota,
                                                          asistencia=recordantiguo.asistencia,
                                                          sinasistencia=recordantiguo.sinasistencia,
                                                          fecha=recordantiguo.fecha,
                                                          noaplica=recordantiguo.noaplica,
                                                          aprobada=recordantiguo.aprobada,
                                                          convalidacion=recordantiguo.convalidacion,
                                                          pendiente=recordantiguo.pendiente,
                                                          creditos=equivalencia.asignaturamallasalto.creditos,
                                                          horas=equivalencia.asignaturamallasalto.horas,
                                                          valida=recordantiguo.valida,
                                                          validapromedio=recordantiguo.validapromedio,
                                                          observaciones=observaciones,
                                                          homologada=homologada,
                                                          materiaregular=recordantiguo.materiaregular,
                                                          materiacurso=None,
                                                          completonota=recordantiguo.completonota,
                                                          completoasistencia=recordantiguo.completoasistencia,
                                                          fechainicio=recordantiguo.fechainicio,
                                                          fechafin=recordantiguo.fechafin,
                                                          suficiencia=recordantiguo.suficiencia,
                                                          asignaturamallahistorico_id=recordantiguo.asignaturamalla.id,
                                                          reverso=False)
                            recordnuevo.save()
                            print(u"Crea nuevo record %s" % recordnuevo)


                        elif RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                            asignaturamalla=equivalencia.asignaturamallasalto):
                            recordnuevo = RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                         asignaturamalla=equivalencia.asignaturamallasalto)[0]
                            recordnuevo.matriculas = recordantiguo.matriculas
                            recordnuevo.asignaturamalla = equivalencia.asignaturamallasalto
                            recordnuevo.asignatura = equivalencia.asignaturamallasalto.asignatura
                            recordnuevo.asignaturaold = recordantiguo.asignaturamalla.asignatura
                            recordnuevo.nota = recordantiguo.nota
                            recordnuevo.asistencia = recordantiguo.asistencia
                            recordnuevo.sinasistencia = recordantiguo.sinasistencia
                            recordnuevo.fecha = recordantiguo.fecha
                            recordnuevo.noaplica = recordantiguo.noaplica
                            recordnuevo.aprobada = recordantiguo.aprobada
                            recordnuevo.convalidacion = recordantiguo.convalidacion
                            recordnuevo.pendiente = recordantiguo.pendiente
                            recordnuevo.creditos = equivalencia.asignaturamallasalto.creditos
                            recordnuevo.horas = equivalencia.asignaturamallasalto.horas
                            recordnuevo.valida = recordantiguo.valida
                            recordnuevo.validapromedio = recordantiguo.validapromedio
                            recordnuevo.observaciones = observaciones
                            recordnuevo.homologada = homologada
                            recordnuevo.materiaregular = recordantiguo.materiaregular
                            recordnuevo.materiacurso = None
                            recordnuevo.completonota = recordantiguo.completonota
                            recordnuevo.completoasistencia = recordantiguo.completoasistencia
                            recordnuevo.fechainicio = recordantiguo.fechainicio
                            recordnuevo.fechafin = recordantiguo.fechafin
                            recordnuevo.suficiencia = recordantiguo.suficiencia
                            recordnuevo.asignaturamallahistorico = recordantiguo.asignaturamalla
                            recordnuevo.reverso = False
                            recordnuevo.save()

                        if recordnuevo:
                            historicos = HistoricoRecordAcademico.objects.filter(status=True,
                                                                                 recordacademico=recordantiguo).update(recordacademico=recordnuevo,
                                                                                                                       creditos=recordnuevo.creditos,
                                                                                                                       horas=recordnuevo.horas,
                                                                                                                       homologada=recordnuevo.homologada)
                            respaldo = RespaldoRecordAcademico.objects.filter(status=True,recordacademicooriginal=recordantiguo)



                            if equivalencia.asignaturamallasalto_id in [11018,11021,11020,10990,10991,11005,10993,10997,11007]:
                                if not practicaspp or not horasvinculacion:
                                    if recordnuevo.aprobada:
                                        profesor = None
                                        if recordnuevo.materiaregular:
                                            profesor = recordnuevo.materiaregular.profesor_principal()
                                        elif recordnuevo.materiacurso:
                                            profesor = recordnuevo.materiaregular.profesor()
                                        if equivalencia.asignaturamallasalto_id == 11018:

                                            itinerarioprimero = ItinerariosMalla.objects.get(status=True, malla_id=mallaantigua_id, nivel_id=1)
                                            iprimeronuevo = ItinerariosMalla.objects.get(status=True, malla_id=mallanueva_id, nivel_id=1)


                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerarioprimero).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerarioprimero).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerarioprimero).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=iprimeronuevo.horas_practicas,
                                                                                                         nivelmalla=iprimeronuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=iprimeronuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerarioprimero).update(
                                                itinerario=iprimeronuevo)

                                        if equivalencia.asignaturamallasalto_id == 11021:
                                            itinerariosegundo = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=2)
                                            isegundonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=2)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerariosegundo).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerariosegundo).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerariosegundo).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=isegundonuevo.horas_practicas,
                                                                                                         nivelmalla=isegundonuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=isegundonuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerariosegundo).update(
                                                itinerario=isegundonuevo)

                                        if equivalencia.asignaturamallasalto_id == 11020:
                                            itinerariotercero = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=3)
                                            iterceronuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=3)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerariotercero).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerariotercero).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerariotercero).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=iterceronuevo.horas_practicas,
                                                                                                         nivelmalla=iterceronuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=iterceronuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerariotercero).update(
                                                itinerario=iterceronuevo)

                                        if equivalencia.asignaturamallasalto_id == 10990:
                                            itinerariocuarto = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=4)
                                            icuartonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=4)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerariocuarto).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerariocuarto).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerariocuarto).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=icuartonuevo.horas_practicas,
                                                                                                         nivelmalla=icuartonuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=icuartonuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerariocuarto).update(
                                                itinerario=icuartonuevo)

                                        if equivalencia.asignaturamallasalto_id == 10991:
                                            itinerarioquinto = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=5)
                                            iquintonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=5)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerarioquinto).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerarioquinto).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerarioquinto).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=iquintonuevo.horas_practicas,
                                                                                                         nivelmalla=iquintonuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=iquintonuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerarioquinto).update(
                                                itinerario=iquintonuevo)

                                        if equivalencia.asignaturamallasalto_id == 11005:
                                            itinerariosexto = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=6)
                                            isextonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=6)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerariosexto).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerariosexto).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerariosexto).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=isextonuevo.horas_practicas,
                                                                                                         nivelmalla=isextonuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=isextonuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerariosexto).update(
                                                itinerario=isextonuevo)

                                        if equivalencia.asignaturamallasalto_id == 10993:
                                            ####################################### VINCULACION ###################################################
                                            if not horasvinculacion:
                                                if inscripcion.numero_horas_proyectos_vinculacion() < 80:
                                                    horasfalta = 80 - inscripcion.numero_horas_proyectos_vinculacion()
                                                    vinculacion = ParticipantesMatrices(status=True,
                                                                                        matrizevidencia_id=2,
                                                                                        proyecto_id=601,
                                                                                        inscripcion=inscripcion,
                                                                                        horas=horasfalta,
                                                                                        registrohorasdesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                        registrohorashasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                        estado=1
                                                                                        )
                                                    vinculacion.save()
                                            ######################################################################################################
                                            itinerarioseptimo = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=7)
                                            iseptimonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=7)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerarioseptimo).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerarioseptimo).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerarioseptimo).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=iseptimonuevo.horas_practicas,
                                                                                                         nivelmalla=iseptimonuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=iseptimonuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerarioseptimo).update(
                                                itinerario=iseptimonuevo)

                                        if equivalencia.asignaturamallasalto_id == 10997:
                                            itinerariooctavo = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=8)
                                            ioctavonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=8)
                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerariooctavo).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerariooctavo).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerariooctavo).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=ioctavonuevo.horas_practicas,
                                                                                                         nivelmalla=ioctavonuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=ioctavonuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerariooctavo).update(
                                                itinerario=ioctavonuevo)

                                        if equivalencia.asignaturamallasalto_id == 11007:
                                            ####################################### VINCULACION ###################################################
                                            if not horasvinculacion:
                                                if inscripcion.numero_horas_proyectos_vinculacion() < 160:
                                                    horasfalta = 160 - inscripcion.numero_horas_proyectos_vinculacion()
                                                    vinculacion = ParticipantesMatrices(status=True,
                                                                                        matrizevidencia_id=2,
                                                                                        proyecto_id=601,
                                                                                        inscripcion=inscripcion,
                                                                                        horas=horasfalta,
                                                                                        registrohorasdesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                        registrohorashasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                        estado=1
                                                                                        )
                                                    vinculacion.save()
                                            ######################################################################################################
                                            itinerarionoveno = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=9)
                                            inovenonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=9)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           estadosolicitud__in=[
                                                                                                               1, 2, 4,
                                                                                                               5, 6],
                                                                                                           itinerariomalla=itinerarionoveno).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerarionoveno).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                           inscripcion=inscripcion,
                                                                                                           actividad__itinerariomalla=itinerarionoveno).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(status=True,
                                                                                                         inscripcion=inscripcion,
                                                                                                         fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         numerohora=inovenonuevo.horas_practicas,
                                                                                                         nivelmalla=inovenonuevo.nivel,
                                                                                                         tiposolicitud=1,
                                                                                                         estadosolicitud=2,
                                                                                                         tipo=1,
                                                                                                         itinerariomalla=inovenonuevo,
                                                                                                         supervisor=profesor,
                                                                                                         tutorunemi=profesor,
                                                                                                         fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         tipoinstitucion=1,
                                                                                                         sectoreconomico=6,
                                                                                                         empresaempleadora_id=3,
                                                                                                         culminada=True,
                                                                                                         fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                                         lugarpractica_id=2,
                                                                                                         observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                                                                         )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(status=True,
                                                                                                          itinerario=itinerarionoveno).update(
                                                itinerario=inovenonuevo)


                            if not respaldo.exists():
                                respaldorecord = RespaldoRecordAcademico(
                                    recordacademicooriginal=recordantiguo,
                                    recordacademiconuevo=recordnuevo
                                )
                                respaldorecord.save()
                            else:
                                respaldorecord = respaldo[0]
                                respaldorecord.recordacademiconuevo = recordnuevo
                                respaldorecord.save()
                            print(u"Record actualizado %s" % recordnuevo)

                    else:
                        hojadestino.write(fila, 3, "NO ENCONTRO RECORD ANTIGUO %s" % equivalencia.asignaturamalla)
                        fila += 1

                practicasppf = inscripcion.numero_horas_practicas_pre_profesionales()
                hojadestino.write(fila, 3, practicasppf, fuentenormal)
                horasvinculacionf = inscripcion.numero_horas_proyectos_vinculacion()
                hojadestino.write(fila, 4, horasvinculacionf, fuentenormal)

                fila += 1

                time.sleep(1)

            lin += 1

        libdestino.save(output_folder + libre_origen)
        print(output_folder + libre_origen)
        print("Proceso finalizado. . .")

    except Exception as ex:
        transaction.set_rollback(True)
        print('error: %s' % ex)
        hojadestino.write(fila, 3, str(ex))
        fila += 1



def homologacion_psicologia1():
    try:
        libre_origen = '/homologacion_psico_2_1.xls'
        fuentecabecera = easyxf(
            'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
        fuentenormal = easyxf(
            'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')

        output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
        # liborigen = xlrd.open_workbook(output_folder + libre_origen)
        libdestino = xlwt.Workbook()
        hojadestino = libdestino.add_sheet('Sheet1')
        fil = 0
        columnas = [(u"CEDULA", 6000),
                    (u"APELLIDOS Y NOMBRES", 6000),
                    (u"HORAS PRACTICAS", 6000),
                    (u"HORAS VINCULACION", 6000),
                    (u"OBSERVACIÓN", 6000)
                    ]
        for col_num in range(len(columnas)):
            hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
            hojadestino.col(col_num).width = columnas[col_num][1]
        fila = 1
        titulacion = 0

        lin = 0
        # miarchivo = openpyxl.load_workbook("CEDULA_COMUNICACION_final.xlsx")
        miarchivo = openpyxl.load_workbook("psiclinea.xlsx")
        # miarchivo = openpyxl.load_workbook("prueba_2.xlsx")

        ws = miarchivo.get_sheet_by_name("primero")
        worksheet = ws
        c = 0
        cont = 0
        periodo_id = 224
        carrera_id = 132
        mallaantigua_id = 204
        mallanueva_id = 479

        for row in worksheet.iter_rows(min_row=0):
            if lin >= 0:
                currentValues, cadena = [], ''
                for cell in row:
                    cadena += str(cell.value) + ' '
                    currentValues.append(str(cell.value))
                identificacion = currentValues[0]

                if not identificacion:
                    break

                with transaction.atomic():
                    matricula = Matricula.objects.filter(status=True, inscripcion__carrera__id=carrera_id,
                                                         inscripcion__persona__cedula=identificacion).first()
                    cont += 1
                    matricula.pasoayuda = True
                    matricula.save()
                    print(u"%s - %s" % (matricula, cont))
                    inscripcion = matricula.inscripcion
                    hojadestino.write(fila, 0, matricula.inscripcion.persona.identificacion(), fuentenormal)
                    hojadestino.write(fila, 1, matricula.inscripcion.persona.nombre_completo(), fuentenormal)
                    hojadestino.write(fila, 2, matricula.nivel.periodo.nombre, fuentenormal)

                    practicaspp = haber_cumplido_horas_creditos_practicas_preprofesionales(inscripcion.id)
                    horasvinculacion = haber_cumplido_horas_creditos_vinculacion(inscripcion.id)

                    if InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                       malla_id=mallaantigua_id).exists():
                        imantigua = \
                            InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                            malla_id=mallaantigua_id)[0]
                        imantigua.status = False
                        imantigua.save()
                        print(u"Desactiva antigua inscripcion -----------------------------")

                    if not InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                           malla_id=mallanueva_id).exists():
                        imnueva = InscripcionMalla(inscripcion=inscripcion, malla_id=mallanueva_id)
                        imnueva.save()
                        print(u"Crea nueva inscripcion -----------------------------")

                    equivalencias = TablaEquivalenciaAsignaturas.objects.filter(status=True,
                                                                                asignaturamalla__malla_id=mallaantigua_id).order_by(
                        'asignaturamallasalto__nivelmalla__orden')

                    cont_asig_vinculacion_aprobadas = 0
                    horasfalta = 0
                    fechainicioitinerario = None
                    fechafinitinerario = None

                    temporal = []
                    for e in equivalencias:
                        temporal.append(
                            [inscripcion.recordacademico_set.filter(status=True,
                                                                    asignaturamalla=e.asignaturamalla).first(),
                             e])

                    for t in temporal:
                        recordantiguo, equivalencia = t

                        old, new = equivalencia.asignaturamalla, equivalencia.asignaturamallasalto
                        print(f"Nueva - {equivalencia.asignaturamallasalto}")

                        if recordantiguo:
                            print(f"Anterior - {equivalencia.asignaturamalla}")
                            print(f"Record antiguo: {recordantiguo}")

                            recordantiguo.status = False
                            recordantiguo.save(update_asignaturamalla=False)

                            recordnuevo, homologada = None, False
                            if equivalencia.asignaturamallasalto_id in (10646, 10649, 10654, 10639):
                                observaciones = recordantiguo.observaciones + " Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15"
                                homologada = True
                            else:
                                observaciones = recordantiguo.observaciones + " Migración con base a RESOLUCIÓN CGA-SO-9-2023-NO15 Y RESOLUCIÓN OCS-SO-18-2023-NO14"

                            if recordnuevo := inscripcion.recordacademico_set.filter(asignaturamalla=new,
                                                                                     status=True).first():
                                recordnuevo.matriculas = recordantiguo.matriculas
                                recordnuevo.asignaturamalla = equivalencia.asignaturamallasalto
                                recordnuevo.asignatura = equivalencia.asignaturamallasalto.asignatura
                                recordnuevo.asignaturaold = recordantiguo.asignaturamalla.asignatura
                                recordnuevo.nota = recordantiguo.nota
                                recordnuevo.asistencia = recordantiguo.asistencia
                                recordnuevo.sinasistencia = recordantiguo.sinasistencia
                                recordnuevo.fecha = recordantiguo.fecha
                                recordnuevo.noaplica = recordantiguo.noaplica
                                recordnuevo.aprobada = recordantiguo.aprobada
                                recordnuevo.convalidacion = recordantiguo.convalidacion
                                recordnuevo.pendiente = recordantiguo.pendiente
                                recordnuevo.creditos = equivalencia.asignaturamallasalto.creditos
                                recordnuevo.horas = equivalencia.asignaturamallasalto.horas
                                recordnuevo.valida = recordantiguo.valida
                                recordnuevo.validapromedio = recordantiguo.validapromedio
                                recordnuevo.observaciones = observaciones
                                recordnuevo.homologada = homologada
                                recordnuevo.materiaregular = recordantiguo.materiaregular
                                recordnuevo.materiacurso = None
                                recordnuevo.completonota = recordantiguo.completonota
                                recordnuevo.completoasistencia = recordantiguo.completoasistencia
                                recordnuevo.fechainicio = recordantiguo.fechainicio
                                recordnuevo.fechafin = recordantiguo.fechafin
                                recordnuevo.suficiencia = recordantiguo.suficiencia
                                recordnuevo.asignaturamallahistorico = recordantiguo.asignaturamalla
                                recordnuevo.reverso = False
                                recordnuevo.save()
                            else:
                                recordnuevo = RecordAcademico(inscripcion=inscripcion,
                                                              matriculas=recordantiguo.matriculas,
                                                              asignaturamalla=equivalencia.asignaturamallasalto,
                                                              asignatura=equivalencia.asignaturamallasalto.asignatura,
                                                              asignaturaold_id=recordantiguo.asignatura.id,
                                                              nota=recordantiguo.nota,
                                                              asistencia=recordantiguo.asistencia,
                                                              sinasistencia=recordantiguo.sinasistencia,
                                                              fecha=recordantiguo.fecha,
                                                              noaplica=recordantiguo.noaplica,
                                                              aprobada=recordantiguo.aprobada,
                                                              convalidacion=recordantiguo.convalidacion,
                                                              pendiente=recordantiguo.pendiente,
                                                              creditos=equivalencia.asignaturamallasalto.creditos,
                                                              horas=equivalencia.asignaturamallasalto.horas,
                                                              valida=recordantiguo.valida,
                                                              validapromedio=recordantiguo.validapromedio,
                                                              observaciones=observaciones,
                                                              homologada=homologada,
                                                              materiaregular=recordantiguo.materiaregular,
                                                              materiacurso=None,
                                                              completonota=recordantiguo.completonota,
                                                              completoasistencia=recordantiguo.completoasistencia,
                                                              fechainicio=recordantiguo.fechainicio,
                                                              fechafin=recordantiguo.fechafin,
                                                              suficiencia=recordantiguo.suficiencia,
                                                              asignaturamallahistorico_id=recordantiguo.asignaturamalla.id,
                                                              reverso=False)
                                recordnuevo.save()
                                print(u"Crea nuevo record %s" % recordnuevo)

                            if recordnuevo:
                                historicos = HistoricoRecordAcademico.objects.filter(status=True,
                                                                                     recordacademico=recordantiguo).update(
                                    recordacademico=recordnuevo, creditos=recordnuevo.creditos, horas=recordnuevo.horas,
                                    homologada=recordnuevo.homologada)
                                respaldo = RespaldoRecordAcademico.objects.filter(status=True,
                                                                                  recordacademicooriginal=recordantiguo)

                                if not respaldo.exists():
                                    respaldorecord = RespaldoRecordAcademico(recordacademicooriginal=recordantiguo,
                                                                             recordacademiconuevo=recordnuevo)
                                    respaldorecord.save()
                                else:
                                    respaldorecord = respaldo[0]
                                    respaldorecord.recordacademiconuevo = recordnuevo
                                    respaldorecord.save()

                                if equivalencia.asignaturamallasalto_id in [10646, 10649, 10654]:
                                    if not practicaspp:
                                        if recordnuevo.aprobada:
                                            profesor = None
                                            if recordnuevo.materiaregular:
                                                profesor = recordnuevo.materiaregular.profesor_principal()
                                            elif recordnuevo.materiacurso:
                                                profesor = recordnuevo.materiaregular.profesor()
                                            if equivalencia.asignaturamallasalto_id == 10646:
                                                itinerariosexto = ItinerariosMalla.objects.get(status=True,
                                                                                               malla_id=mallaantigua_id,
                                                                                               nivel_id=6)

                                                isextonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallanueva_id,
                                                                                           nivel_id=6)

                                                practica = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud__in=[
                                                        1, 2, 4,
                                                        5, 6],
                                                    itinerariomalla=itinerariosexto).exists()
                                                practicarechazada = False
                                                if not practica:
                                                    practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        estadosolicitud=3,
                                                        itinerariomalla=itinerariosexto).exists()
                                                if not practica or practicarechazada:
                                                    if not PracticasPreprofesionalesInscripcion.objects.filter(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            actividad__itinerariomalla=itinerariosexto).exists():
                                                        nuevapractica = PracticasPreprofesionalesInscripcion(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            numerohora=isextonuevo.horas_practicas,
                                                            nivelmalla=isextonuevo.nivel,
                                                            tiposolicitud=1,
                                                            estadosolicitud=2,
                                                            tipo=1,
                                                            itinerariomalla=isextonuevo,
                                                            supervisor=profesor,
                                                            tutorunemi=profesor,
                                                            fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            tipoinstitucion=1,
                                                            sectoreconomico=6,
                                                            empresaempleadora_id=3,
                                                            culminada=True,
                                                            fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            lugarpractica_id=2,
                                                            observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                        )
                                                        nuevapractica.save()

                                                ItinerariosActividadDetalleDistributivoCarrera.objects.filter(
                                                    status=True,
                                                    itinerario=itinerariosexto).update(
                                                    itinerario=isextonuevo)

                                            if equivalencia.asignaturamallasalto_id == 10649:
                                                itinerarioseptimo = ItinerariosMalla.objects.get(status=True,
                                                                                                 malla_id=mallaantigua_id,
                                                                                                 nivel_id=7)
                                                iseptimonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                             malla_id=mallanueva_id,
                                                                                             nivel_id=7)

                                                practica = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud__in=[
                                                        1, 2, 4,
                                                        5, 6],
                                                    itinerariomalla=itinerarioseptimo).exists()
                                                practicarechazada = False
                                                if not practica:
                                                    practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        estadosolicitud=3,
                                                        itinerariomalla=itinerarioseptimo).exists()

                                                if not practica or practicarechazada:
                                                    if not PracticasPreprofesionalesInscripcion.objects.filter(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            actividad__itinerariomalla=itinerarioseptimo).exists():
                                                        nuevapractica = PracticasPreprofesionalesInscripcion(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                            fechahasta=recordnuevo.materiaregular.nivel.periodo.fin,
                                                            numerohora=iseptimonuevo.horas_practicas,
                                                            nivelmalla=iseptimonuevo.nivel,
                                                            tiposolicitud=1,
                                                            estadosolicitud=2,
                                                            tipo=1,
                                                            itinerariomalla=iseptimonuevo,
                                                            supervisor=profesor,
                                                            tutorunemi=profesor,
                                                            fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                            tipoinstitucion=1,
                                                            sectoreconomico=6,
                                                            empresaempleadora_id=3,
                                                            culminada=True,
                                                            fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                            lugarpractica_id=2,
                                                            observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                        )
                                                        nuevapractica.save()

                                                ItinerariosActividadDetalleDistributivoCarrera.objects.filter(
                                                    status=True,
                                                    itinerario=itinerarioseptimo).update(
                                                    itinerario=iseptimonuevo)

                                            if equivalencia.asignaturamallasalto_id == 10654:
                                                itinerariooctavo = ItinerariosMalla.objects.get(status=True,
                                                                                                malla_id=mallaantigua_id,
                                                                                                nivel_id=8)
                                                ioctavonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                            malla_id=mallanueva_id,
                                                                                            nivel_id=8)

                                                practica = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud__in=[
                                                        1, 2, 4,
                                                        5, 6],
                                                    itinerariomalla=itinerariooctavo).exists()
                                                practicarechazada = False
                                                if not practica:
                                                    practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        estadosolicitud=3,
                                                        itinerariomalla=itinerariooctavo).exists()

                                                if not practica or practicarechazada:
                                                    if not PracticasPreprofesionalesInscripcion.objects.filter(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            actividad__itinerariomalla=itinerariooctavo).exists():
                                                        nuevapractica = PracticasPreprofesionalesInscripcion(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            numerohora=ioctavonuevo.horas_practicas,
                                                            nivelmalla=ioctavonuevo.nivel,
                                                            tiposolicitud=1,
                                                            estadosolicitud=2,
                                                            tipo=1,
                                                            itinerariomalla=ioctavonuevo,
                                                            supervisor=profesor,
                                                            tutorunemi=profesor,
                                                            fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            tipoinstitucion=1,
                                                            sectoreconomico=6,
                                                            empresaempleadora_id=3,
                                                            culminada=True,
                                                            fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            lugarpractica_id=2,
                                                            observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                        )
                                                        nuevapractica.save()
                                                ItinerariosActividadDetalleDistributivoCarrera.objects.filter(
                                                    status=True,
                                                    itinerario=itinerariooctavo).update(
                                                    itinerario=ioctavonuevo)

                                if equivalencia.asignaturamallasalto_id in [10639, 10654]:
                                    if not horasvinculacion:
                                        if recordnuevo.aprobada:
                                            if equivalencia.asignaturamallasalto_id == 10639:
                                                cont_asig_vinculacion_aprobadas += 1

                                            if equivalencia.asignaturamallasalto_id == 10654:
                                                cont_asig_vinculacion_aprobadas += 1
                                            fechainicioitinerario = recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date()
                                            fechafinitinerario = recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date()

                                print(u"Record actualizado %s" % recordnuevo)
                        else:
                            hojadestino.write(fila, 5, "NO ENCONTRO RECORD ANTIGUO %s" % equivalencia.asignaturamalla)
                            fila += 1

                    if cont_asig_vinculacion_aprobadas != 0:
                        if cont_asig_vinculacion_aprobadas == 1:
                            horasfalta = 144
                        if cont_asig_vinculacion_aprobadas == 2:
                            horasfalta = 160
                        horasfalta = horasfalta - inscripcion.numero_horas_proyectos_vinculacion()
                        vinculacion = ParticipantesMatrices(status=True,
                                                            matrizevidencia_id=2,
                                                            proyecto_id=601,
                                                            inscripcion=inscripcion,
                                                            horas=horasfalta,
                                                            registrohorasdesde=fechainicioitinerario,
                                                            registrohorashasta=fechafinitinerario,
                                                            estado=1
                                                            )
                        vinculacion.save()

                    practicasppf = inscripcion.numero_horas_practicas_pre_profesionales()
                    hojadestino.write(fila, 3, practicasppf, fuentenormal)
                    horasvinculacionf = inscripcion.numero_horas_proyectos_vinculacion()
                    hojadestino.write(fila, 4, horasvinculacionf, fuentenormal)

                    fila += 1

                    time.sleep(3)



            lin += 1

        libdestino.save(output_folder + libre_origen)
        print(output_folder + libre_origen)
        print("Proceso finalizado. . .")

    except Exception as ex:
        noti = Notificacion(titulo='Error',
                            cuerpo='Ha ocurrido un error {} - Error en la linea {}'.format(
                                ex, sys.exc_info()[-1].tb_lineno),
                            destinatario_id=29898, url="",
                            prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                            tipo=2, en_proceso=False, error=True)
        noti.save()
        transaction.set_rollback(True)
        print('error: %s' % ex)
        hojadestino.write(fila, 3, str(ex))
        fila += 1

def homologacion_psicologia2():
    try:
        libre_origen = '/homologacion_psico_2_2.xls'
        fuentecabecera = easyxf(
            'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
        fuentenormal = easyxf(
            'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')

        output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
        # liborigen = xlrd.open_workbook(output_folder + libre_origen)
        libdestino = xlwt.Workbook()
        hojadestino = libdestino.add_sheet('Sheet1')
        fil = 0
        columnas = [(u"CEDULA", 6000),
                    (u"APELLIDOS Y NOMBRES", 6000),
                    (u"HORAS PRACTICAS", 6000),
                    (u"HORAS VINCULACION", 6000),
                    (u"OBSERVACIÓN", 6000)
                    ]
        for col_num in range(len(columnas)):
            hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
            hojadestino.col(col_num).width = columnas[col_num][1]
        fila = 1
        titulacion = 0

        lin = 0
        # miarchivo = openpyxl.load_workbook("CEDULA_COMUNICACION_final.xlsx")
        miarchivo = openpyxl.load_workbook("psiclinea.xlsx")
        # miarchivo = openpyxl.load_workbook("prueba_2.xlsx")

        ws = miarchivo.get_sheet_by_name("segundo")
        worksheet = ws
        c = 0
        cont = 0
        periodo_id = 224
        carrera_id = 132
        mallaantigua_id = 204
        mallanueva_id = 479

        for row in worksheet.iter_rows(min_row=0):
            if lin >= 0:
                currentValues, cadena = [], ''
                for cell in row:
                    cadena += str(cell.value) + ' '
                    currentValues.append(str(cell.value))
                identificacion = currentValues[0]

                if not identificacion:
                    break

                with transaction.atomic():
                    matricula = Matricula.objects.filter(status=True, inscripcion__carrera__id=carrera_id,
                                                         inscripcion__persona__cedula=identificacion).first()
                    cont += 1
                    matricula.pasoayuda = True
                    matricula.save()
                    print(u"%s - %s" % (matricula, cont))
                    inscripcion = matricula.inscripcion
                    hojadestino.write(fila, 0, matricula.inscripcion.persona.identificacion(), fuentenormal)
                    hojadestino.write(fila, 1, matricula.inscripcion.persona.nombre_completo(), fuentenormal)
                    hojadestino.write(fila, 2, matricula.nivel.periodo.nombre, fuentenormal)

                    practicaspp = haber_cumplido_horas_creditos_practicas_preprofesionales(inscripcion.id)
                    horasvinculacion = haber_cumplido_horas_creditos_vinculacion(inscripcion.id)

                    if InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                       malla_id=mallaantigua_id).exists():
                        imantigua = \
                            InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                            malla_id=mallaantigua_id)[0]
                        imantigua.status = False
                        imantigua.save()
                        print(u"Desactiva antigua inscripcion -----------------------------")

                    if not InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                           malla_id=mallanueva_id).exists():
                        imnueva = InscripcionMalla(inscripcion=inscripcion, malla_id=mallanueva_id)
                        imnueva.save()
                        print(u"Crea nueva inscripcion -----------------------------")

                    equivalencias = TablaEquivalenciaAsignaturas.objects.filter(status=True,
                                                                                asignaturamalla__malla_id=mallaantigua_id).order_by(
                        'asignaturamallasalto__nivelmalla__orden')

                    cont_asig_vinculacion_aprobadas = 0
                    horasfalta = 0
                    fechainicioitinerario = None
                    fechafinitinerario = None

                    temporal = []
                    for e in equivalencias:
                        temporal.append(
                            [inscripcion.recordacademico_set.filter(status=True,
                                                                    asignaturamalla=e.asignaturamalla).first(),
                             e])

                    for t in temporal:
                        recordantiguo, equivalencia = t

                        old, new = equivalencia.asignaturamalla, equivalencia.asignaturamallasalto
                        print(f"Nueva - {equivalencia.asignaturamallasalto}")

                        if recordantiguo:
                            print(f"Anterior - {equivalencia.asignaturamalla}")
                            print(f"Record antiguo: {recordantiguo}")

                            recordantiguo.status = False
                            recordantiguo.save(update_asignaturamalla=False)

                            recordnuevo, homologada = None, False
                            if equivalencia.asignaturamallasalto_id in (10646, 10649, 10654, 10639):
                                observaciones = recordantiguo.observaciones + " Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15"
                                homologada = True
                            else:
                                observaciones = recordantiguo.observaciones + " Migración con base a RESOLUCIÓN CGA-SO-9-2023-NO15 Y RESOLUCIÓN OCS-SO-18-2023-NO14"

                            if recordnuevo := inscripcion.recordacademico_set.filter(asignaturamalla=new,
                                                                                     status=True).first():
                                recordnuevo.matriculas = recordantiguo.matriculas
                                recordnuevo.asignaturamalla = equivalencia.asignaturamallasalto
                                recordnuevo.asignatura = equivalencia.asignaturamallasalto.asignatura
                                recordnuevo.asignaturaold = recordantiguo.asignaturamalla.asignatura
                                recordnuevo.nota = recordantiguo.nota
                                recordnuevo.asistencia = recordantiguo.asistencia
                                recordnuevo.sinasistencia = recordantiguo.sinasistencia
                                recordnuevo.fecha = recordantiguo.fecha
                                recordnuevo.noaplica = recordantiguo.noaplica
                                recordnuevo.aprobada = recordantiguo.aprobada
                                recordnuevo.convalidacion = recordantiguo.convalidacion
                                recordnuevo.pendiente = recordantiguo.pendiente
                                recordnuevo.creditos = equivalencia.asignaturamallasalto.creditos
                                recordnuevo.horas = equivalencia.asignaturamallasalto.horas
                                recordnuevo.valida = recordantiguo.valida
                                recordnuevo.validapromedio = recordantiguo.validapromedio
                                recordnuevo.observaciones = observaciones
                                recordnuevo.homologada = homologada
                                recordnuevo.materiaregular = recordantiguo.materiaregular
                                recordnuevo.materiacurso = None
                                recordnuevo.completonota = recordantiguo.completonota
                                recordnuevo.completoasistencia = recordantiguo.completoasistencia
                                recordnuevo.fechainicio = recordantiguo.fechainicio
                                recordnuevo.fechafin = recordantiguo.fechafin
                                recordnuevo.suficiencia = recordantiguo.suficiencia
                                recordnuevo.asignaturamallahistorico = recordantiguo.asignaturamalla
                                recordnuevo.reverso = False
                                recordnuevo.save()
                            else:
                                recordnuevo = RecordAcademico(inscripcion=inscripcion,
                                                              matriculas=recordantiguo.matriculas,
                                                              asignaturamalla=equivalencia.asignaturamallasalto,
                                                              asignatura=equivalencia.asignaturamallasalto.asignatura,
                                                              asignaturaold_id=recordantiguo.asignatura.id,
                                                              nota=recordantiguo.nota,
                                                              asistencia=recordantiguo.asistencia,
                                                              sinasistencia=recordantiguo.sinasistencia,
                                                              fecha=recordantiguo.fecha,
                                                              noaplica=recordantiguo.noaplica,
                                                              aprobada=recordantiguo.aprobada,
                                                              convalidacion=recordantiguo.convalidacion,
                                                              pendiente=recordantiguo.pendiente,
                                                              creditos=equivalencia.asignaturamallasalto.creditos,
                                                              horas=equivalencia.asignaturamallasalto.horas,
                                                              valida=recordantiguo.valida,
                                                              validapromedio=recordantiguo.validapromedio,
                                                              observaciones=observaciones,
                                                              homologada=homologada,
                                                              materiaregular=recordantiguo.materiaregular,
                                                              materiacurso=None,
                                                              completonota=recordantiguo.completonota,
                                                              completoasistencia=recordantiguo.completoasistencia,
                                                              fechainicio=recordantiguo.fechainicio,
                                                              fechafin=recordantiguo.fechafin,
                                                              suficiencia=recordantiguo.suficiencia,
                                                              asignaturamallahistorico_id=recordantiguo.asignaturamalla.id,
                                                              reverso=False)
                                recordnuevo.save()
                                print(u"Crea nuevo record %s" % recordnuevo)

                            if recordnuevo:
                                historicos = HistoricoRecordAcademico.objects.filter(status=True,
                                                                                     recordacademico=recordantiguo).update(
                                    recordacademico=recordnuevo, creditos=recordnuevo.creditos, horas=recordnuevo.horas,
                                    homologada=recordnuevo.homologada)
                                respaldo = RespaldoRecordAcademico.objects.filter(status=True,
                                                                                  recordacademicooriginal=recordantiguo)

                                if not respaldo.exists():
                                    respaldorecord = RespaldoRecordAcademico(recordacademicooriginal=recordantiguo,
                                                                             recordacademiconuevo=recordnuevo)
                                    respaldorecord.save()
                                else:
                                    respaldorecord = respaldo[0]
                                    respaldorecord.recordacademiconuevo = recordnuevo
                                    respaldorecord.save()

                                if equivalencia.asignaturamallasalto_id in [10646, 10649, 10654]:
                                    if not practicaspp:
                                        if recordnuevo.aprobada:
                                            profesor = None
                                            if recordnuevo.materiaregular:
                                                profesor = recordnuevo.materiaregular.profesor_principal()
                                            elif recordnuevo.materiacurso:
                                                profesor = recordnuevo.materiaregular.profesor()
                                            if equivalencia.asignaturamallasalto_id == 10646:
                                                itinerariosexto = ItinerariosMalla.objects.get(status=True,
                                                                                               malla_id=mallaantigua_id,
                                                                                               nivel_id=6)

                                                isextonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallanueva_id,
                                                                                           nivel_id=6)

                                                practica = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud__in=[
                                                        1, 2, 4,
                                                        5, 6],
                                                    itinerariomalla=itinerariosexto).exists()
                                                practicarechazada = False
                                                if not practica:
                                                    practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        estadosolicitud=3,
                                                        itinerariomalla=itinerariosexto).exists()
                                                if not practica or practicarechazada:
                                                    if not PracticasPreprofesionalesInscripcion.objects.filter(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            actividad__itinerariomalla=itinerariosexto).exists():
                                                        nuevapractica = PracticasPreprofesionalesInscripcion(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            numerohora=isextonuevo.horas_practicas,
                                                            nivelmalla=isextonuevo.nivel,
                                                            tiposolicitud=1,
                                                            estadosolicitud=2,
                                                            tipo=1,
                                                            itinerariomalla=isextonuevo,
                                                            supervisor=profesor,
                                                            tutorunemi=profesor,
                                                            fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            tipoinstitucion=1,
                                                            sectoreconomico=6,
                                                            empresaempleadora_id=3,
                                                            culminada=True,
                                                            fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            lugarpractica_id=2,
                                                            observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                        )
                                                        nuevapractica.save()

                                                ItinerariosActividadDetalleDistributivoCarrera.objects.filter(
                                                    status=True,
                                                    itinerario=itinerariosexto).update(
                                                    itinerario=isextonuevo)

                                            if equivalencia.asignaturamallasalto_id == 10649:
                                                itinerarioseptimo = ItinerariosMalla.objects.get(status=True,
                                                                                                 malla_id=mallaantigua_id,
                                                                                                 nivel_id=7)
                                                iseptimonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                             malla_id=mallanueva_id,
                                                                                             nivel_id=7)

                                                practica = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud__in=[
                                                        1, 2, 4,
                                                        5, 6],
                                                    itinerariomalla=itinerarioseptimo).exists()
                                                practicarechazada = False
                                                if not practica:
                                                    practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        estadosolicitud=3,
                                                        itinerariomalla=itinerarioseptimo).exists()

                                                if not practica or practicarechazada:
                                                    if not PracticasPreprofesionalesInscripcion.objects.filter(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            actividad__itinerariomalla=itinerarioseptimo).exists():
                                                        nuevapractica = PracticasPreprofesionalesInscripcion(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                            fechahasta=recordnuevo.materiaregular.nivel.periodo.fin,
                                                            numerohora=iseptimonuevo.horas_practicas,
                                                            nivelmalla=iseptimonuevo.nivel,
                                                            tiposolicitud=1,
                                                            estadosolicitud=2,
                                                            tipo=1,
                                                            itinerariomalla=iseptimonuevo,
                                                            supervisor=profesor,
                                                            tutorunemi=profesor,
                                                            fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                            tipoinstitucion=1,
                                                            sectoreconomico=6,
                                                            empresaempleadora_id=3,
                                                            culminada=True,
                                                            fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                            lugarpractica_id=2,
                                                            observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                        )
                                                        nuevapractica.save()

                                                ItinerariosActividadDetalleDistributivoCarrera.objects.filter(
                                                    status=True,
                                                    itinerario=itinerarioseptimo).update(
                                                    itinerario=iseptimonuevo)

                                            if equivalencia.asignaturamallasalto_id == 10654:
                                                itinerariooctavo = ItinerariosMalla.objects.get(status=True,
                                                                                                malla_id=mallaantigua_id,
                                                                                                nivel_id=8)
                                                ioctavonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                            malla_id=mallanueva_id,
                                                                                            nivel_id=8)

                                                practica = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud__in=[
                                                        1, 2, 4,
                                                        5, 6],
                                                    itinerariomalla=itinerariooctavo).exists()
                                                practicarechazada = False
                                                if not practica:
                                                    practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        estadosolicitud=3,
                                                        itinerariomalla=itinerariooctavo).exists()

                                                if not practica or practicarechazada:
                                                    if not PracticasPreprofesionalesInscripcion.objects.filter(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            actividad__itinerariomalla=itinerariooctavo).exists():
                                                        nuevapractica = PracticasPreprofesionalesInscripcion(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            numerohora=ioctavonuevo.horas_practicas,
                                                            nivelmalla=ioctavonuevo.nivel,
                                                            tiposolicitud=1,
                                                            estadosolicitud=2,
                                                            tipo=1,
                                                            itinerariomalla=ioctavonuevo,
                                                            supervisor=profesor,
                                                            tutorunemi=profesor,
                                                            fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            tipoinstitucion=1,
                                                            sectoreconomico=6,
                                                            empresaempleadora_id=3,
                                                            culminada=True,
                                                            fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            lugarpractica_id=2,
                                                            observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                        )
                                                        nuevapractica.save()
                                                ItinerariosActividadDetalleDistributivoCarrera.objects.filter(
                                                    status=True,
                                                    itinerario=itinerariooctavo).update(
                                                    itinerario=ioctavonuevo)

                                if equivalencia.asignaturamallasalto_id in [10639, 10654]:
                                    if not horasvinculacion:
                                        if recordnuevo.aprobada:
                                            if equivalencia.asignaturamallasalto_id == 10639:
                                                cont_asig_vinculacion_aprobadas += 1

                                            if equivalencia.asignaturamallasalto_id == 10654:
                                                cont_asig_vinculacion_aprobadas += 1
                                            fechainicioitinerario = recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date()
                                            fechafinitinerario = recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date()

                                print(u"Record actualizado %s" % recordnuevo)
                        else:
                            hojadestino.write(fila, 5, "NO ENCONTRO RECORD ANTIGUO %s" % equivalencia.asignaturamalla)
                            fila += 1

                    if cont_asig_vinculacion_aprobadas != 0:
                        if cont_asig_vinculacion_aprobadas == 1:
                            horasfalta = 144
                        if cont_asig_vinculacion_aprobadas == 2:
                            horasfalta = 160
                        horasfalta = horasfalta - inscripcion.numero_horas_proyectos_vinculacion()
                        vinculacion = ParticipantesMatrices(status=True,
                                                            matrizevidencia_id=2,
                                                            proyecto_id=601,
                                                            inscripcion=inscripcion,
                                                            horas=horasfalta,
                                                            registrohorasdesde=fechainicioitinerario,
                                                            registrohorashasta=fechafinitinerario,
                                                            estado=1
                                                            )
                        vinculacion.save()

                    practicasppf = inscripcion.numero_horas_practicas_pre_profesionales()
                    hojadestino.write(fila, 3, practicasppf, fuentenormal)
                    horasvinculacionf = inscripcion.numero_horas_proyectos_vinculacion()
                    hojadestino.write(fila, 4, horasvinculacionf, fuentenormal)

                    fila += 1

                    time.sleep(3)


            lin += 1

        libdestino.save(output_folder + libre_origen)
        print(output_folder + libre_origen)
        print("Proceso finalizado. . .")

    except Exception as ex:
        noti = Notificacion(titulo='Error',
                            cuerpo='Ha ocurrido un error {} - Error en la linea {}'.format(
                                ex, sys.exc_info()[-1].tb_lineno),
                            destinatario_id=29898, url="",
                            prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                            tipo=2, en_proceso=False, error=True)
        noti.save()
        transaction.set_rollback(True)
        print('error: %s' % ex)
        hojadestino.write(fila, 3, str(ex))
        fila += 1

def homologacion_psicologia3():
    try:
        libre_origen = '/homologacion_psico_2_3.xls'
        fuentecabecera = easyxf(
            'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
        fuentenormal = easyxf(
            'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')

        output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
        # liborigen = xlrd.open_workbook(output_folder + libre_origen)
        libdestino = xlwt.Workbook()
        hojadestino = libdestino.add_sheet('Sheet1')
        fil = 0
        columnas = [(u"CEDULA", 6000),
                    (u"APELLIDOS Y NOMBRES", 6000),
                    (u"HORAS PRACTICAS", 6000),
                    (u"HORAS VINCULACION", 6000),
                    (u"OBSERVACIÓN", 6000)
                    ]
        for col_num in range(len(columnas)):
            hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
            hojadestino.col(col_num).width = columnas[col_num][1]
        fila = 1
        titulacion = 0

        lin = 0
        # miarchivo = openpyxl.load_workbook("CEDULA_COMUNICACION_final.xlsx")
        miarchivo = openpyxl.load_workbook("psiclinea.xlsx")
        # miarchivo = openpyxl.load_workbook("prueba_2.xlsx")

        ws = miarchivo.get_sheet_by_name("tercero")
        worksheet = ws
        c = 0
        cont = 0
        periodo_id = 224
        carrera_id = 132
        mallaantigua_id = 204
        mallanueva_id = 479

        for row in worksheet.iter_rows(min_row=0):
            if lin >= 0:
                currentValues, cadena = [], ''
                for cell in row:
                    cadena += str(cell.value) + ' '
                    currentValues.append(str(cell.value))
                identificacion = currentValues[0]

                if not identificacion:
                    break

                with transaction.atomic():
                    matricula = Matricula.objects.filter(status=True, inscripcion__carrera__id=carrera_id,
                                                         inscripcion__persona__cedula=identificacion).first()
                    cont += 1
                    matricula.pasoayuda = True
                    matricula.save()
                    print(u"%s - %s" % (matricula, cont))
                    inscripcion = matricula.inscripcion
                    hojadestino.write(fila, 0, matricula.inscripcion.persona.identificacion(), fuentenormal)
                    hojadestino.write(fila, 1, matricula.inscripcion.persona.nombre_completo(), fuentenormal)
                    hojadestino.write(fila, 2, matricula.nivel.periodo.nombre, fuentenormal)

                    practicaspp = haber_cumplido_horas_creditos_practicas_preprofesionales(inscripcion.id)
                    horasvinculacion = haber_cumplido_horas_creditos_vinculacion(inscripcion.id)

                    if InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                       malla_id=mallaantigua_id).exists():
                        imantigua = \
                            InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                            malla_id=mallaantigua_id)[0]
                        imantigua.status = False
                        imantigua.save()
                        print(u"Desactiva antigua inscripcion -----------------------------")

                    if not InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                           malla_id=mallanueva_id).exists():
                        imnueva = InscripcionMalla(inscripcion=inscripcion, malla_id=mallanueva_id)
                        imnueva.save()
                        print(u"Crea nueva inscripcion -----------------------------")

                    equivalencias = TablaEquivalenciaAsignaturas.objects.filter(status=True,
                                                                                asignaturamalla__malla_id=mallaantigua_id).order_by(
                        'asignaturamallasalto__nivelmalla__orden')

                    cont_asig_vinculacion_aprobadas = 0
                    horasfalta = 0
                    fechainicioitinerario = None
                    fechafinitinerario = None

                    temporal = []
                    for e in equivalencias:
                        temporal.append(
                            [inscripcion.recordacademico_set.filter(status=True,
                                                                    asignaturamalla=e.asignaturamalla).first(),
                             e])

                    for t in temporal:
                        recordantiguo, equivalencia = t

                        old, new = equivalencia.asignaturamalla, equivalencia.asignaturamallasalto
                        print(f"Nueva - {equivalencia.asignaturamallasalto}")

                        if recordantiguo:
                            print(f"Anterior - {equivalencia.asignaturamalla}")
                            print(f"Record antiguo: {recordantiguo}")

                            recordantiguo.status = False
                            recordantiguo.save(update_asignaturamalla=False)

                            recordnuevo, homologada = None, False
                            if equivalencia.asignaturamallasalto_id in (10646, 10649, 10654, 10639):
                                observaciones = recordantiguo.observaciones + " Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15"
                                homologada = True
                            else:
                                observaciones = recordantiguo.observaciones + " Migración con base a RESOLUCIÓN CGA-SO-9-2023-NO15 Y RESOLUCIÓN OCS-SO-18-2023-NO14"

                            if recordnuevo := inscripcion.recordacademico_set.filter(asignaturamalla=new,
                                                                                     status=True).first():
                                recordnuevo.matriculas = recordantiguo.matriculas
                                recordnuevo.asignaturamalla = equivalencia.asignaturamallasalto
                                recordnuevo.asignatura = equivalencia.asignaturamallasalto.asignatura
                                recordnuevo.asignaturaold = recordantiguo.asignaturamalla.asignatura
                                recordnuevo.nota = recordantiguo.nota
                                recordnuevo.asistencia = recordantiguo.asistencia
                                recordnuevo.sinasistencia = recordantiguo.sinasistencia
                                recordnuevo.fecha = recordantiguo.fecha
                                recordnuevo.noaplica = recordantiguo.noaplica
                                recordnuevo.aprobada = recordantiguo.aprobada
                                recordnuevo.convalidacion = recordantiguo.convalidacion
                                recordnuevo.pendiente = recordantiguo.pendiente
                                recordnuevo.creditos = equivalencia.asignaturamallasalto.creditos
                                recordnuevo.horas = equivalencia.asignaturamallasalto.horas
                                recordnuevo.valida = recordantiguo.valida
                                recordnuevo.validapromedio = recordantiguo.validapromedio
                                recordnuevo.observaciones = observaciones
                                recordnuevo.homologada = homologada
                                recordnuevo.materiaregular = recordantiguo.materiaregular
                                recordnuevo.materiacurso = None
                                recordnuevo.completonota = recordantiguo.completonota
                                recordnuevo.completoasistencia = recordantiguo.completoasistencia
                                recordnuevo.fechainicio = recordantiguo.fechainicio
                                recordnuevo.fechafin = recordantiguo.fechafin
                                recordnuevo.suficiencia = recordantiguo.suficiencia
                                recordnuevo.asignaturamallahistorico = recordantiguo.asignaturamalla
                                recordnuevo.reverso = False
                                recordnuevo.save()
                            else:
                                recordnuevo = RecordAcademico(inscripcion=inscripcion,
                                                              matriculas=recordantiguo.matriculas,
                                                              asignaturamalla=equivalencia.asignaturamallasalto,
                                                              asignatura=equivalencia.asignaturamallasalto.asignatura,
                                                              asignaturaold_id=recordantiguo.asignatura.id,
                                                              nota=recordantiguo.nota,
                                                              asistencia=recordantiguo.asistencia,
                                                              sinasistencia=recordantiguo.sinasistencia,
                                                              fecha=recordantiguo.fecha,
                                                              noaplica=recordantiguo.noaplica,
                                                              aprobada=recordantiguo.aprobada,
                                                              convalidacion=recordantiguo.convalidacion,
                                                              pendiente=recordantiguo.pendiente,
                                                              creditos=equivalencia.asignaturamallasalto.creditos,
                                                              horas=equivalencia.asignaturamallasalto.horas,
                                                              valida=recordantiguo.valida,
                                                              validapromedio=recordantiguo.validapromedio,
                                                              observaciones=observaciones,
                                                              homologada=homologada,
                                                              materiaregular=recordantiguo.materiaregular,
                                                              materiacurso=None,
                                                              completonota=recordantiguo.completonota,
                                                              completoasistencia=recordantiguo.completoasistencia,
                                                              fechainicio=recordantiguo.fechainicio,
                                                              fechafin=recordantiguo.fechafin,
                                                              suficiencia=recordantiguo.suficiencia,
                                                              asignaturamallahistorico_id=recordantiguo.asignaturamalla.id,
                                                              reverso=False)
                                recordnuevo.save()
                                print(u"Crea nuevo record %s" % recordnuevo)

                            if recordnuevo:
                                historicos = HistoricoRecordAcademico.objects.filter(status=True,
                                                                                     recordacademico=recordantiguo).update(
                                    recordacademico=recordnuevo, creditos=recordnuevo.creditos, horas=recordnuevo.horas,
                                    homologada=recordnuevo.homologada)
                                respaldo = RespaldoRecordAcademico.objects.filter(status=True,
                                                                                  recordacademicooriginal=recordantiguo)

                                if not respaldo.exists():
                                    respaldorecord = RespaldoRecordAcademico(recordacademicooriginal=recordantiguo,
                                                                             recordacademiconuevo=recordnuevo)
                                    respaldorecord.save()
                                else:
                                    respaldorecord = respaldo[0]
                                    respaldorecord.recordacademiconuevo = recordnuevo
                                    respaldorecord.save()

                                if equivalencia.asignaturamallasalto_id in [10646, 10649, 10654]:
                                    if not practicaspp:
                                        if recordnuevo.aprobada:
                                            profesor = None
                                            if recordnuevo.materiaregular:
                                                profesor = recordnuevo.materiaregular.profesor_principal()
                                            elif recordnuevo.materiacurso:
                                                profesor = recordnuevo.materiaregular.profesor()
                                            if equivalencia.asignaturamallasalto_id == 10646:
                                                itinerariosexto = ItinerariosMalla.objects.get(status=True,
                                                                                               malla_id=mallaantigua_id,
                                                                                               nivel_id=6)

                                                isextonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallanueva_id,
                                                                                           nivel_id=6)

                                                practica = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud__in=[
                                                        1, 2, 4,
                                                        5, 6],
                                                    itinerariomalla=itinerariosexto).exists()
                                                practicarechazada = False
                                                if not practica:
                                                    practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        estadosolicitud=3,
                                                        itinerariomalla=itinerariosexto).exists()
                                                if not practica or practicarechazada:
                                                    if not PracticasPreprofesionalesInscripcion.objects.filter(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            actividad__itinerariomalla=itinerariosexto).exists():
                                                        nuevapractica = PracticasPreprofesionalesInscripcion(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            numerohora=isextonuevo.horas_practicas,
                                                            nivelmalla=isextonuevo.nivel,
                                                            tiposolicitud=1,
                                                            estadosolicitud=2,
                                                            tipo=1,
                                                            itinerariomalla=isextonuevo,
                                                            supervisor=profesor,
                                                            tutorunemi=profesor,
                                                            fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            tipoinstitucion=1,
                                                            sectoreconomico=6,
                                                            empresaempleadora_id=3,
                                                            culminada=True,
                                                            fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            lugarpractica_id=2,
                                                            observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                        )
                                                        nuevapractica.save()

                                                ItinerariosActividadDetalleDistributivoCarrera.objects.filter(
                                                    status=True,
                                                    itinerario=itinerariosexto).update(
                                                    itinerario=isextonuevo)

                                            if equivalencia.asignaturamallasalto_id == 10649:
                                                itinerarioseptimo = ItinerariosMalla.objects.get(status=True,
                                                                                                 malla_id=mallaantigua_id,
                                                                                                 nivel_id=7)
                                                iseptimonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                             malla_id=mallanueva_id,
                                                                                             nivel_id=7)

                                                practica = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud__in=[
                                                        1, 2, 4,
                                                        5, 6],
                                                    itinerariomalla=itinerarioseptimo).exists()
                                                practicarechazada = False
                                                if not practica:
                                                    practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        estadosolicitud=3,
                                                        itinerariomalla=itinerarioseptimo).exists()

                                                if not practica or practicarechazada:
                                                    if not PracticasPreprofesionalesInscripcion.objects.filter(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            actividad__itinerariomalla=itinerarioseptimo).exists():
                                                        nuevapractica = PracticasPreprofesionalesInscripcion(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                            fechahasta=recordnuevo.materiaregular.nivel.periodo.fin,
                                                            numerohora=iseptimonuevo.horas_practicas,
                                                            nivelmalla=iseptimonuevo.nivel,
                                                            tiposolicitud=1,
                                                            estadosolicitud=2,
                                                            tipo=1,
                                                            itinerariomalla=iseptimonuevo,
                                                            supervisor=profesor,
                                                            tutorunemi=profesor,
                                                            fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                            tipoinstitucion=1,
                                                            sectoreconomico=6,
                                                            empresaempleadora_id=3,
                                                            culminada=True,
                                                            fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                            lugarpractica_id=2,
                                                            observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                        )
                                                        nuevapractica.save()

                                                ItinerariosActividadDetalleDistributivoCarrera.objects.filter(
                                                    status=True,
                                                    itinerario=itinerarioseptimo).update(
                                                    itinerario=iseptimonuevo)

                                            if equivalencia.asignaturamallasalto_id == 10654:
                                                itinerariooctavo = ItinerariosMalla.objects.get(status=True,
                                                                                                malla_id=mallaantigua_id,
                                                                                                nivel_id=8)
                                                ioctavonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                            malla_id=mallanueva_id,
                                                                                            nivel_id=8)

                                                practica = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud__in=[
                                                        1, 2, 4,
                                                        5, 6],
                                                    itinerariomalla=itinerariooctavo).exists()
                                                practicarechazada = False
                                                if not practica:
                                                    practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        estadosolicitud=3,
                                                        itinerariomalla=itinerariooctavo).exists()

                                                if not practica or practicarechazada:
                                                    if not PracticasPreprofesionalesInscripcion.objects.filter(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            actividad__itinerariomalla=itinerariooctavo).exists():
                                                        nuevapractica = PracticasPreprofesionalesInscripcion(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            numerohora=ioctavonuevo.horas_practicas,
                                                            nivelmalla=ioctavonuevo.nivel,
                                                            tiposolicitud=1,
                                                            estadosolicitud=2,
                                                            tipo=1,
                                                            itinerariomalla=ioctavonuevo,
                                                            supervisor=profesor,
                                                            tutorunemi=profesor,
                                                            fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            tipoinstitucion=1,
                                                            sectoreconomico=6,
                                                            empresaempleadora_id=3,
                                                            culminada=True,
                                                            fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            lugarpractica_id=2,
                                                            observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                        )
                                                        nuevapractica.save()
                                                ItinerariosActividadDetalleDistributivoCarrera.objects.filter(
                                                    status=True,
                                                    itinerario=itinerariooctavo).update(
                                                    itinerario=ioctavonuevo)

                                if equivalencia.asignaturamallasalto_id in [10639, 10654]:
                                    if not horasvinculacion:
                                        if recordnuevo.aprobada:
                                            if equivalencia.asignaturamallasalto_id == 10639:
                                                cont_asig_vinculacion_aprobadas += 1

                                            if equivalencia.asignaturamallasalto_id == 10654:
                                                cont_asig_vinculacion_aprobadas += 1
                                            fechainicioitinerario = recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date()
                                            fechafinitinerario = recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date()

                                print(u"Record actualizado %s" % recordnuevo)
                        else:
                            hojadestino.write(fila, 5, "NO ENCONTRO RECORD ANTIGUO %s" % equivalencia.asignaturamalla)
                            fila += 1

                    if cont_asig_vinculacion_aprobadas != 0:
                        if cont_asig_vinculacion_aprobadas == 1:
                            horasfalta = 144
                        if cont_asig_vinculacion_aprobadas == 2:
                            horasfalta = 160
                        horasfalta = horasfalta - inscripcion.numero_horas_proyectos_vinculacion()
                        vinculacion = ParticipantesMatrices(status=True,
                                                            matrizevidencia_id=2,
                                                            proyecto_id=601,
                                                            inscripcion=inscripcion,
                                                            horas=horasfalta,
                                                            registrohorasdesde=fechainicioitinerario,
                                                            registrohorashasta=fechafinitinerario,
                                                            estado=1
                                                            )
                        vinculacion.save()

                    practicasppf = inscripcion.numero_horas_practicas_pre_profesionales()
                    hojadestino.write(fila, 3, practicasppf, fuentenormal)
                    horasvinculacionf = inscripcion.numero_horas_proyectos_vinculacion()
                    hojadestino.write(fila, 4, horasvinculacionf, fuentenormal)

                    fila += 1

                    time.sleep(3)


            lin += 1

        libdestino.save(output_folder + libre_origen)
        print(output_folder + libre_origen)
        print("Proceso finalizado. . .")

    except Exception as ex:
        noti = Notificacion(titulo='Error',
                            cuerpo='Ha ocurrido un error {} - Error en la linea {}'.format(
                                ex, sys.exc_info()[-1].tb_lineno),
                            destinatario_id=29898, url="",
                            prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                            tipo=2, en_proceso=False, error=True)
        noti.save()
        transaction.set_rollback(True)
        print('error: %s' % ex)
        hojadestino.write(fila, 3, str(ex))
        fila += 1

def homologacion_psicologia4():
    try:
        libre_origen = '/homologacion_psico_2_4.xls'
        fuentecabecera = easyxf(
            'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
        fuentenormal = easyxf(
            'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')

        output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
        # liborigen = xlrd.open_workbook(output_folder + libre_origen)
        libdestino = xlwt.Workbook()
        hojadestino = libdestino.add_sheet('Sheet1')
        fil = 0
        columnas = [(u"CEDULA", 6000),
                    (u"APELLIDOS Y NOMBRES", 6000),
                    (u"HORAS PRACTICAS", 6000),
                    (u"HORAS VINCULACION", 6000),
                    (u"OBSERVACIÓN", 6000)
                    ]
        for col_num in range(len(columnas)):
            hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
            hojadestino.col(col_num).width = columnas[col_num][1]
        fila = 1
        titulacion = 0

        lin = 0
        # miarchivo = openpyxl.load_workbook("CEDULA_COMUNICACION_final.xlsx")
        miarchivo = openpyxl.load_workbook("psiclinea.xlsx")
        # miarchivo = openpyxl.load_workbook("prueba_2.xlsx")

        ws = miarchivo.get_sheet_by_name("cuarto")
        worksheet = ws
        c = 0
        cont = 0
        periodo_id = 224
        carrera_id = 132
        mallaantigua_id = 204
        mallanueva_id = 479

        for row in worksheet.iter_rows(min_row=0):
            if lin >= 0:
                currentValues, cadena = [], ''
                for cell in row:
                    cadena += str(cell.value) + ' '
                    currentValues.append(str(cell.value))
                identificacion = currentValues[0]

                if not identificacion:
                    break

                with transaction.atomic():
                    matricula = Matricula.objects.filter(status=True, inscripcion__carrera__id=carrera_id,
                                                         inscripcion__persona__cedula=identificacion).first()
                    cont += 1
                    matricula.pasoayuda = True
                    matricula.save()
                    print(u"%s - %s" % (matricula, cont))
                    inscripcion = matricula.inscripcion
                    hojadestino.write(fila, 0, matricula.inscripcion.persona.identificacion(), fuentenormal)
                    hojadestino.write(fila, 1, matricula.inscripcion.persona.nombre_completo(), fuentenormal)
                    hojadestino.write(fila, 2, matricula.nivel.periodo.nombre, fuentenormal)

                    practicaspp = haber_cumplido_horas_creditos_practicas_preprofesionales(inscripcion.id)
                    horasvinculacion = haber_cumplido_horas_creditos_vinculacion(inscripcion.id)

                    if InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                       malla_id=mallaantigua_id).exists():
                        imantigua = \
                            InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                            malla_id=mallaantigua_id)[0]
                        imantigua.status = False
                        imantigua.save()
                        print(u"Desactiva antigua inscripcion -----------------------------")

                    if not InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                           malla_id=mallanueva_id).exists():
                        imnueva = InscripcionMalla(inscripcion=inscripcion, malla_id=mallanueva_id)
                        imnueva.save()
                        print(u"Crea nueva inscripcion -----------------------------")

                    equivalencias = TablaEquivalenciaAsignaturas.objects.filter(status=True,
                                                                                asignaturamalla__malla_id=mallaantigua_id).order_by(
                        'asignaturamallasalto__nivelmalla__orden')

                    cont_asig_vinculacion_aprobadas = 0
                    horasfalta = 0
                    fechainicioitinerario = None
                    fechafinitinerario = None

                    temporal = []
                    for e in equivalencias:
                        temporal.append(
                            [inscripcion.recordacademico_set.filter(status=True,
                                                                    asignaturamalla=e.asignaturamalla).first(),
                             e])

                    for t in temporal:
                        recordantiguo, equivalencia = t

                        old, new = equivalencia.asignaturamalla, equivalencia.asignaturamallasalto
                        print(f"Nueva - {equivalencia.asignaturamallasalto}")

                        if recordantiguo:
                            print(f"Anterior - {equivalencia.asignaturamalla}")
                            print(f"Record antiguo: {recordantiguo}")

                            recordantiguo.status = False
                            recordantiguo.save(update_asignaturamalla=False)

                            recordnuevo, homologada = None, False
                            if equivalencia.asignaturamallasalto_id in (10646, 10649, 10654, 10639):
                                observaciones = recordantiguo.observaciones + " Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15"
                                homologada = True
                            else:
                                observaciones = recordantiguo.observaciones + " Migración con base a RESOLUCIÓN CGA-SO-9-2023-NO15 Y RESOLUCIÓN OCS-SO-18-2023-NO14"

                            if recordnuevo := inscripcion.recordacademico_set.filter(asignaturamalla=new,
                                                                                     status=True).first():
                                recordnuevo.matriculas = recordantiguo.matriculas
                                recordnuevo.asignaturamalla = equivalencia.asignaturamallasalto
                                recordnuevo.asignatura = equivalencia.asignaturamallasalto.asignatura
                                recordnuevo.asignaturaold = recordantiguo.asignaturamalla.asignatura
                                recordnuevo.nota = recordantiguo.nota
                                recordnuevo.asistencia = recordantiguo.asistencia
                                recordnuevo.sinasistencia = recordantiguo.sinasistencia
                                recordnuevo.fecha = recordantiguo.fecha
                                recordnuevo.noaplica = recordantiguo.noaplica
                                recordnuevo.aprobada = recordantiguo.aprobada
                                recordnuevo.convalidacion = recordantiguo.convalidacion
                                recordnuevo.pendiente = recordantiguo.pendiente
                                recordnuevo.creditos = equivalencia.asignaturamallasalto.creditos
                                recordnuevo.horas = equivalencia.asignaturamallasalto.horas
                                recordnuevo.valida = recordantiguo.valida
                                recordnuevo.validapromedio = recordantiguo.validapromedio
                                recordnuevo.observaciones = observaciones
                                recordnuevo.homologada = homologada
                                recordnuevo.materiaregular = recordantiguo.materiaregular
                                recordnuevo.materiacurso = None
                                recordnuevo.completonota = recordantiguo.completonota
                                recordnuevo.completoasistencia = recordantiguo.completoasistencia
                                recordnuevo.fechainicio = recordantiguo.fechainicio
                                recordnuevo.fechafin = recordantiguo.fechafin
                                recordnuevo.suficiencia = recordantiguo.suficiencia
                                recordnuevo.asignaturamallahistorico = recordantiguo.asignaturamalla
                                recordnuevo.reverso = False
                                recordnuevo.save()
                            else:
                                recordnuevo = RecordAcademico(inscripcion=inscripcion,
                                                              matriculas=recordantiguo.matriculas,
                                                              asignaturamalla=equivalencia.asignaturamallasalto,
                                                              asignatura=equivalencia.asignaturamallasalto.asignatura,
                                                              asignaturaold_id=recordantiguo.asignatura.id,
                                                              nota=recordantiguo.nota,
                                                              asistencia=recordantiguo.asistencia,
                                                              sinasistencia=recordantiguo.sinasistencia,
                                                              fecha=recordantiguo.fecha,
                                                              noaplica=recordantiguo.noaplica,
                                                              aprobada=recordantiguo.aprobada,
                                                              convalidacion=recordantiguo.convalidacion,
                                                              pendiente=recordantiguo.pendiente,
                                                              creditos=equivalencia.asignaturamallasalto.creditos,
                                                              horas=equivalencia.asignaturamallasalto.horas,
                                                              valida=recordantiguo.valida,
                                                              validapromedio=recordantiguo.validapromedio,
                                                              observaciones=observaciones,
                                                              homologada=homologada,
                                                              materiaregular=recordantiguo.materiaregular,
                                                              materiacurso=None,
                                                              completonota=recordantiguo.completonota,
                                                              completoasistencia=recordantiguo.completoasistencia,
                                                              fechainicio=recordantiguo.fechainicio,
                                                              fechafin=recordantiguo.fechafin,
                                                              suficiencia=recordantiguo.suficiencia,
                                                              asignaturamallahistorico_id=recordantiguo.asignaturamalla.id,
                                                              reverso=False)
                                recordnuevo.save()
                                print(u"Crea nuevo record %s" % recordnuevo)

                            if recordnuevo:
                                historicos = HistoricoRecordAcademico.objects.filter(status=True,
                                                                                     recordacademico=recordantiguo).update(
                                    recordacademico=recordnuevo, creditos=recordnuevo.creditos, horas=recordnuevo.horas,
                                    homologada=recordnuevo.homologada)
                                respaldo = RespaldoRecordAcademico.objects.filter(status=True,
                                                                                  recordacademicooriginal=recordantiguo)

                                if not respaldo.exists():
                                    respaldorecord = RespaldoRecordAcademico(recordacademicooriginal=recordantiguo,
                                                                             recordacademiconuevo=recordnuevo)
                                    respaldorecord.save()
                                else:
                                    respaldorecord = respaldo[0]
                                    respaldorecord.recordacademiconuevo = recordnuevo
                                    respaldorecord.save()

                                if equivalencia.asignaturamallasalto_id in [10646, 10649, 10654]:
                                    if not practicaspp:
                                        if recordnuevo.aprobada:
                                            profesor = None
                                            if recordnuevo.materiaregular:
                                                profesor = recordnuevo.materiaregular.profesor_principal()
                                            elif recordnuevo.materiacurso:
                                                profesor = recordnuevo.materiaregular.profesor()
                                            if equivalencia.asignaturamallasalto_id == 10646:
                                                itinerariosexto = ItinerariosMalla.objects.get(status=True,
                                                                                               malla_id=mallaantigua_id,
                                                                                               nivel_id=6)

                                                isextonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallanueva_id,
                                                                                           nivel_id=6)

                                                practica = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud__in=[
                                                        1, 2, 4,
                                                        5, 6],
                                                    itinerariomalla=itinerariosexto).exists()
                                                practicarechazada = False
                                                if not practica:
                                                    practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        estadosolicitud=3,
                                                        itinerariomalla=itinerariosexto).exists()
                                                if not practica or practicarechazada:
                                                    if not PracticasPreprofesionalesInscripcion.objects.filter(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            actividad__itinerariomalla=itinerariosexto).exists():
                                                        nuevapractica = PracticasPreprofesionalesInscripcion(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            numerohora=isextonuevo.horas_practicas,
                                                            nivelmalla=isextonuevo.nivel,
                                                            tiposolicitud=1,
                                                            estadosolicitud=2,
                                                            tipo=1,
                                                            itinerariomalla=isextonuevo,
                                                            supervisor=profesor,
                                                            tutorunemi=profesor,
                                                            fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            tipoinstitucion=1,
                                                            sectoreconomico=6,
                                                            empresaempleadora_id=3,
                                                            culminada=True,
                                                            fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            lugarpractica_id=2,
                                                            observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                        )
                                                        nuevapractica.save()

                                                ItinerariosActividadDetalleDistributivoCarrera.objects.filter(
                                                    status=True,
                                                    itinerario=itinerariosexto).update(
                                                    itinerario=isextonuevo)

                                            if equivalencia.asignaturamallasalto_id == 10649:
                                                itinerarioseptimo = ItinerariosMalla.objects.get(status=True,
                                                                                                 malla_id=mallaantigua_id,
                                                                                                 nivel_id=7)
                                                iseptimonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                             malla_id=mallanueva_id,
                                                                                             nivel_id=7)

                                                practica = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud__in=[
                                                        1, 2, 4,
                                                        5, 6],
                                                    itinerariomalla=itinerarioseptimo).exists()
                                                practicarechazada = False
                                                if not practica:
                                                    practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        estadosolicitud=3,
                                                        itinerariomalla=itinerarioseptimo).exists()

                                                if not practica or practicarechazada:
                                                    if not PracticasPreprofesionalesInscripcion.objects.filter(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            actividad__itinerariomalla=itinerarioseptimo).exists():
                                                        nuevapractica = PracticasPreprofesionalesInscripcion(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                            fechahasta=recordnuevo.materiaregular.nivel.periodo.fin,
                                                            numerohora=iseptimonuevo.horas_practicas,
                                                            nivelmalla=iseptimonuevo.nivel,
                                                            tiposolicitud=1,
                                                            estadosolicitud=2,
                                                            tipo=1,
                                                            itinerariomalla=iseptimonuevo,
                                                            supervisor=profesor,
                                                            tutorunemi=profesor,
                                                            fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                            tipoinstitucion=1,
                                                            sectoreconomico=6,
                                                            empresaempleadora_id=3,
                                                            culminada=True,
                                                            fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                            lugarpractica_id=2,
                                                            observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                        )
                                                        nuevapractica.save()

                                                ItinerariosActividadDetalleDistributivoCarrera.objects.filter(
                                                    status=True,
                                                    itinerario=itinerarioseptimo).update(
                                                    itinerario=iseptimonuevo)

                                            if equivalencia.asignaturamallasalto_id == 10654:
                                                itinerariooctavo = ItinerariosMalla.objects.get(status=True,
                                                                                                malla_id=mallaantigua_id,
                                                                                                nivel_id=8)
                                                ioctavonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                            malla_id=mallanueva_id,
                                                                                            nivel_id=8)

                                                practica = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud__in=[
                                                        1, 2, 4,
                                                        5, 6],
                                                    itinerariomalla=itinerariooctavo).exists()
                                                practicarechazada = False
                                                if not practica:
                                                    practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        estadosolicitud=3,
                                                        itinerariomalla=itinerariooctavo).exists()

                                                if not practica or practicarechazada:
                                                    if not PracticasPreprofesionalesInscripcion.objects.filter(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            actividad__itinerariomalla=itinerariooctavo).exists():
                                                        nuevapractica = PracticasPreprofesionalesInscripcion(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            numerohora=ioctavonuevo.horas_practicas,
                                                            nivelmalla=ioctavonuevo.nivel,
                                                            tiposolicitud=1,
                                                            estadosolicitud=2,
                                                            tipo=1,
                                                            itinerariomalla=ioctavonuevo,
                                                            supervisor=profesor,
                                                            tutorunemi=profesor,
                                                            fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            tipoinstitucion=1,
                                                            sectoreconomico=6,
                                                            empresaempleadora_id=3,
                                                            culminada=True,
                                                            fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            lugarpractica_id=2,
                                                            observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                        )
                                                        nuevapractica.save()
                                                ItinerariosActividadDetalleDistributivoCarrera.objects.filter(
                                                    status=True,
                                                    itinerario=itinerariooctavo).update(
                                                    itinerario=ioctavonuevo)

                                if equivalencia.asignaturamallasalto_id in [10639, 10654]:
                                    if not horasvinculacion:
                                        if recordnuevo.aprobada:
                                            if equivalencia.asignaturamallasalto_id == 10639:
                                                cont_asig_vinculacion_aprobadas += 1

                                            if equivalencia.asignaturamallasalto_id == 10654:
                                                cont_asig_vinculacion_aprobadas += 1
                                            fechainicioitinerario = recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date()
                                            fechafinitinerario = recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date()

                                print(u"Record actualizado %s" % recordnuevo)
                        else:
                            hojadestino.write(fila, 5, "NO ENCONTRO RECORD ANTIGUO %s" % equivalencia.asignaturamalla)
                            fila += 1

                    if cont_asig_vinculacion_aprobadas != 0:
                        if cont_asig_vinculacion_aprobadas == 1:
                            horasfalta = 144
                        if cont_asig_vinculacion_aprobadas == 2:
                            horasfalta = 160
                        horasfalta = horasfalta - inscripcion.numero_horas_proyectos_vinculacion()
                        vinculacion = ParticipantesMatrices(status=True,
                                                            matrizevidencia_id=2,
                                                            proyecto_id=601,
                                                            inscripcion=inscripcion,
                                                            horas=horasfalta,
                                                            registrohorasdesde=fechainicioitinerario,
                                                            registrohorashasta=fechafinitinerario,
                                                            estado=1
                                                            )
                        vinculacion.save()

                    practicasppf = inscripcion.numero_horas_practicas_pre_profesionales()
                    hojadestino.write(fila, 3, practicasppf, fuentenormal)
                    horasvinculacionf = inscripcion.numero_horas_proyectos_vinculacion()
                    hojadestino.write(fila, 4, horasvinculacionf, fuentenormal)

                    fila += 1

                    time.sleep(3)


            lin += 1

        libdestino.save(output_folder + libre_origen)
        print(output_folder + libre_origen)
        print("Proceso finalizado. . .")

    except Exception as ex:
        noti = Notificacion(titulo='Error',
                            cuerpo='Ha ocurrido un error {} - Error en la linea {}'.format(
                                ex, sys.exc_info()[-1].tb_lineno),
                            destinatario_id=29898, url="",
                            prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                            tipo=2, en_proceso=False, error=True)
        noti.save()
        transaction.set_rollback(True)
        print('error: %s' % ex)
        hojadestino.write(fila, 3, str(ex))
        fila += 1

def homologacion_psicologia5():
    try:
        libre_origen = '/homologacion_psico_2_5.xls'
        fuentecabecera = easyxf(
            'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
        fuentenormal = easyxf(
            'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')

        output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
        # liborigen = xlrd.open_workbook(output_folder + libre_origen)
        libdestino = xlwt.Workbook()
        hojadestino = libdestino.add_sheet('Sheet1')
        fil = 0
        columnas = [(u"CEDULA", 6000),
                    (u"APELLIDOS Y NOMBRES", 6000),
                    (u"HORAS PRACTICAS", 6000),
                    (u"HORAS VINCULACION", 6000),
                    (u"OBSERVACIÓN", 6000)
                    ]
        for col_num in range(len(columnas)):
            hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
            hojadestino.col(col_num).width = columnas[col_num][1]
        fila = 1
        titulacion = 0

        lin = 0
        # miarchivo = openpyxl.load_workbook("CEDULA_COMUNICACION_final.xlsx")
        miarchivo = openpyxl.load_workbook("psiclinea.xlsx")
        # miarchivo = openpyxl.load_workbook("prueba_2.xlsx")

        ws = miarchivo.get_sheet_by_name("quinto")
        worksheet = ws
        c = 0
        cont = 0
        periodo_id = 224
        carrera_id = 132
        mallaantigua_id = 204
        mallanueva_id = 479

        for row in worksheet.iter_rows(min_row=0):
            if lin >= 0:
                currentValues, cadena = [], ''
                for cell in row:
                    cadena += str(cell.value) + ' '
                    currentValues.append(str(cell.value))
                identificacion = currentValues[0]

                if not identificacion:
                    break

                matricula = Matricula.objects.filter(status=True, inscripcion__carrera__id=carrera_id,
                                                     inscripcion__persona__cedula=identificacion).first()
                cont += 1
                matricula.pasoayuda = True
                matricula.save()
                print(u"%s - %s" % (matricula, cont))
                inscripcion = matricula.inscripcion
                hojadestino.write(fila, 0, matricula.inscripcion.persona.identificacion(), fuentenormal)
                hojadestino.write(fila, 1, matricula.inscripcion.persona.nombre_completo(), fuentenormal)
                hojadestino.write(fila, 2, matricula.nivel.periodo.nombre, fuentenormal)

                practicaspp = haber_cumplido_horas_creditos_practicas_preprofesionales(inscripcion.id)
                horasvinculacion = haber_cumplido_horas_creditos_vinculacion(inscripcion.id)

                if InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                   malla_id=mallaantigua_id).exists():
                    imantigua = \
                        InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                        malla_id=mallaantigua_id)[0]
                    imantigua.status = False
                    imantigua.save()
                    print(u"Desactiva antigua inscripcion -----------------------------")

                if not InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                       malla_id=mallanueva_id).exists():
                    imnueva = InscripcionMalla(inscripcion=inscripcion, malla_id=mallanueva_id)
                    imnueva.save()
                    print(u"Crea nueva inscripcion -----------------------------")

                equivalencias = TablaEquivalenciaAsignaturas.objects.filter(status=True,
                                                                            asignaturamalla__malla_id=mallaantigua_id).order_by(
                    'asignaturamallasalto__nivelmalla__orden')

                cont_asig_vinculacion_aprobadas = 0
                horasfalta = 0
                fechainicioitinerario = None
                fechafinitinerario = None

                temporal = []
                for e in equivalencias:
                    temporal.append(
                        [inscripcion.recordacademico_set.filter(status=True,
                                                                asignaturamalla=e.asignaturamalla).first(),
                         e])

                for t in temporal:
                    recordantiguo, equivalencia = t

                    old, new = equivalencia.asignaturamalla, equivalencia.asignaturamallasalto
                    print(f"Nueva - {equivalencia.asignaturamallasalto}")

                    if recordantiguo:
                        print(f"Anterior - {equivalencia.asignaturamalla}")
                        print(f"Record antiguo: {recordantiguo}")

                        recordantiguo.status = False
                        recordantiguo.save(update_asignaturamalla=False)

                        recordnuevo, homologada = None, False
                        if equivalencia.asignaturamallasalto_id in (10646, 10649, 10654, 10639):
                            observaciones = recordantiguo.observaciones + " Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15"
                            homologada = True
                        else:
                            observaciones = recordantiguo.observaciones + " Migración con base a RESOLUCIÓN CGA-SO-9-2023-NO15 Y RESOLUCIÓN OCS-SO-18-2023-NO14"

                        if recordnuevo := inscripcion.recordacademico_set.filter(asignaturamalla=new,
                                                                                 status=True).first():
                            recordnuevo.matriculas = recordantiguo.matriculas
                            recordnuevo.asignaturamalla = equivalencia.asignaturamallasalto
                            recordnuevo.asignatura = equivalencia.asignaturamallasalto.asignatura
                            recordnuevo.asignaturaold = recordantiguo.asignaturamalla.asignatura
                            recordnuevo.nota = recordantiguo.nota
                            recordnuevo.asistencia = recordantiguo.asistencia
                            recordnuevo.sinasistencia = recordantiguo.sinasistencia
                            recordnuevo.fecha = recordantiguo.fecha
                            recordnuevo.noaplica = recordantiguo.noaplica
                            recordnuevo.aprobada = recordantiguo.aprobada
                            recordnuevo.convalidacion = recordantiguo.convalidacion
                            recordnuevo.pendiente = recordantiguo.pendiente
                            recordnuevo.creditos = equivalencia.asignaturamallasalto.creditos
                            recordnuevo.horas = equivalencia.asignaturamallasalto.horas
                            recordnuevo.valida = recordantiguo.valida
                            recordnuevo.validapromedio = recordantiguo.validapromedio
                            recordnuevo.observaciones = observaciones
                            recordnuevo.homologada = homologada
                            recordnuevo.materiaregular = recordantiguo.materiaregular
                            recordnuevo.materiacurso = None
                            recordnuevo.completonota = recordantiguo.completonota
                            recordnuevo.completoasistencia = recordantiguo.completoasistencia
                            recordnuevo.fechainicio = recordantiguo.fechainicio
                            recordnuevo.fechafin = recordantiguo.fechafin
                            recordnuevo.suficiencia = recordantiguo.suficiencia
                            recordnuevo.asignaturamallahistorico = recordantiguo.asignaturamalla
                            recordnuevo.reverso = False
                            recordnuevo.save()
                        else:
                            recordnuevo = RecordAcademico(inscripcion=inscripcion,
                                                          matriculas=recordantiguo.matriculas,
                                                          asignaturamalla=equivalencia.asignaturamallasalto,
                                                          asignatura=equivalencia.asignaturamallasalto.asignatura,
                                                          asignaturaold_id=recordantiguo.asignatura.id,
                                                          nota=recordantiguo.nota,
                                                          asistencia=recordantiguo.asistencia,
                                                          sinasistencia=recordantiguo.sinasistencia,
                                                          fecha=recordantiguo.fecha,
                                                          noaplica=recordantiguo.noaplica,
                                                          aprobada=recordantiguo.aprobada,
                                                          convalidacion=recordantiguo.convalidacion,
                                                          pendiente=recordantiguo.pendiente,
                                                          creditos=equivalencia.asignaturamallasalto.creditos,
                                                          horas=equivalencia.asignaturamallasalto.horas,
                                                          valida=recordantiguo.valida,
                                                          validapromedio=recordantiguo.validapromedio,
                                                          observaciones=observaciones,
                                                          homologada=homologada,
                                                          materiaregular=recordantiguo.materiaregular,
                                                          materiacurso=None,
                                                          completonota=recordantiguo.completonota,
                                                          completoasistencia=recordantiguo.completoasistencia,
                                                          fechainicio=recordantiguo.fechainicio,
                                                          fechafin=recordantiguo.fechafin,
                                                          suficiencia=recordantiguo.suficiencia,
                                                          asignaturamallahistorico_id=recordantiguo.asignaturamalla.id,
                                                          reverso=False)
                            recordnuevo.save()
                            print(u"Crea nuevo record %s" % recordnuevo)

                        if recordnuevo:
                            historicos = HistoricoRecordAcademico.objects.filter(status=True,
                                                                                 recordacademico=recordantiguo).update(
                                recordacademico=recordnuevo, creditos=recordnuevo.creditos, horas=recordnuevo.horas,
                                homologada=recordnuevo.homologada)
                            respaldo = RespaldoRecordAcademico.objects.filter(status=True,
                                                                              recordacademicooriginal=recordantiguo)

                            if not respaldo.exists():
                                respaldorecord = RespaldoRecordAcademico(recordacademicooriginal=recordantiguo,
                                                                         recordacademiconuevo=recordnuevo)
                                respaldorecord.save()
                            else:
                                respaldorecord = respaldo[0]
                                respaldorecord.recordacademiconuevo = recordnuevo
                                respaldorecord.save()

                            if equivalencia.asignaturamallasalto_id in [10646, 10649, 10654]:
                                if not practicaspp:
                                    if recordnuevo.aprobada:
                                        profesor = None
                                        if recordnuevo.materiaregular:
                                            profesor = recordnuevo.materiaregular.profesor_principal()
                                        elif recordnuevo.materiacurso:
                                            profesor = recordnuevo.materiaregular.profesor()
                                        if equivalencia.asignaturamallasalto_id == 10646:
                                            itinerariosexto = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=6)

                                            isextonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                       malla_id=mallanueva_id,
                                                                                       nivel_id=6)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(
                                                status=True,
                                                inscripcion=inscripcion,
                                                estadosolicitud__in=[
                                                    1, 2, 4,
                                                    5, 6],
                                                itinerariomalla=itinerariosexto).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerariosexto).exists()
                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        actividad__itinerariomalla=itinerariosexto).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                        fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                        numerohora=isextonuevo.horas_practicas,
                                                        nivelmalla=isextonuevo.nivel,
                                                        tiposolicitud=1,
                                                        estadosolicitud=2,
                                                        tipo=1,
                                                        itinerariomalla=isextonuevo,
                                                        supervisor=profesor,
                                                        tutorunemi=profesor,
                                                        fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                        tipoinstitucion=1,
                                                        sectoreconomico=6,
                                                        empresaempleadora_id=3,
                                                        culminada=True,
                                                        fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                        lugarpractica_id=2,
                                                        observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                    )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(
                                                status=True,
                                                itinerario=itinerariosexto).update(
                                                itinerario=isextonuevo)

                                        if equivalencia.asignaturamallasalto_id == 10649:
                                            itinerarioseptimo = ItinerariosMalla.objects.get(status=True,
                                                                                             malla_id=mallaantigua_id,
                                                                                             nivel_id=7)
                                            iseptimonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                         malla_id=mallanueva_id,
                                                                                         nivel_id=7)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(
                                                status=True,
                                                inscripcion=inscripcion,
                                                estadosolicitud__in=[
                                                    1, 2, 4,
                                                    5, 6],
                                                itinerariomalla=itinerarioseptimo).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerarioseptimo).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        actividad__itinerariomalla=itinerarioseptimo).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                        fechahasta=recordnuevo.materiaregular.nivel.periodo.fin,
                                                        numerohora=iseptimonuevo.horas_practicas,
                                                        nivelmalla=iseptimonuevo.nivel,
                                                        tiposolicitud=1,
                                                        estadosolicitud=2,
                                                        tipo=1,
                                                        itinerariomalla=iseptimonuevo,
                                                        supervisor=profesor,
                                                        tutorunemi=profesor,
                                                        fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                        tipoinstitucion=1,
                                                        sectoreconomico=6,
                                                        empresaempleadora_id=3,
                                                        culminada=True,
                                                        fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                        lugarpractica_id=2,
                                                        observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                    )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(
                                                status=True,
                                                itinerario=itinerarioseptimo).update(
                                                itinerario=iseptimonuevo)

                                        if equivalencia.asignaturamallasalto_id == 10654:
                                            itinerariooctavo = ItinerariosMalla.objects.get(status=True,
                                                                                            malla_id=mallaantigua_id,
                                                                                            nivel_id=8)
                                            ioctavonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=8)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(
                                                status=True,
                                                inscripcion=inscripcion,
                                                estadosolicitud__in=[
                                                    1, 2, 4,
                                                    5, 6],
                                                itinerariomalla=itinerariooctavo).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerariooctavo).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        actividad__itinerariomalla=itinerariooctavo).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                        fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                        numerohora=ioctavonuevo.horas_practicas,
                                                        nivelmalla=ioctavonuevo.nivel,
                                                        tiposolicitud=1,
                                                        estadosolicitud=2,
                                                        tipo=1,
                                                        itinerariomalla=ioctavonuevo,
                                                        supervisor=profesor,
                                                        tutorunemi=profesor,
                                                        fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                        tipoinstitucion=1,
                                                        sectoreconomico=6,
                                                        empresaempleadora_id=3,
                                                        culminada=True,
                                                        fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                        lugarpractica_id=2,
                                                        observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                    )
                                                    nuevapractica.save()
                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(
                                                status=True,
                                                itinerario=itinerariooctavo).update(
                                                itinerario=ioctavonuevo)

                            if equivalencia.asignaturamallasalto_id in [10639, 10654]:
                                if not horasvinculacion:
                                    if recordnuevo.aprobada:
                                        if equivalencia.asignaturamallasalto_id == 10639:
                                            cont_asig_vinculacion_aprobadas += 1

                                        if equivalencia.asignaturamallasalto_id == 10654:
                                            cont_asig_vinculacion_aprobadas += 1
                                        fechainicioitinerario = recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date()
                                        fechafinitinerario = recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date()

                            print(u"Record actualizado %s" % recordnuevo)
                    else:
                        hojadestino.write(fila, 5, "NO ENCONTRO RECORD ANTIGUO %s" % equivalencia.asignaturamalla)
                        fila += 1

                if cont_asig_vinculacion_aprobadas != 0:
                    if cont_asig_vinculacion_aprobadas == 1:
                        horasfalta = 144
                    if cont_asig_vinculacion_aprobadas == 2:
                        horasfalta = 160
                    horasfalta = horasfalta - inscripcion.numero_horas_proyectos_vinculacion()
                    vinculacion = ParticipantesMatrices(status=True,
                                                        matrizevidencia_id=2,
                                                        proyecto_id=601,
                                                        inscripcion=inscripcion,
                                                        horas=horasfalta,
                                                        registrohorasdesde=fechainicioitinerario,
                                                        registrohorashasta=fechafinitinerario,
                                                        estado=1
                                                        )
                    vinculacion.save()

                practicasppf = inscripcion.numero_horas_practicas_pre_profesionales()
                hojadestino.write(fila, 3, practicasppf, fuentenormal)
                horasvinculacionf = inscripcion.numero_horas_proyectos_vinculacion()
                hojadestino.write(fila, 4, horasvinculacionf, fuentenormal)

                fila += 1

                time.sleep(3)

            lin += 1

        libdestino.save(output_folder + libre_origen)
        print(output_folder + libre_origen)
        print("Proceso finalizado. . .")

    except Exception as ex:
        noti = Notificacion(titulo='Error',
                            cuerpo='Ha ocurrido un error {} - Error en la linea {}'.format(
                                ex, sys.exc_info()[-1].tb_lineno),
                            destinatario_id=29898, url="",
                            prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                            tipo=2, en_proceso=False, error=True)
        noti.save()
        transaction.set_rollback(True)
        print('error: %s' % ex)
        hojadestino.write(fila, 3, str(ex))
        fila += 1

def homologacion_psicologia6():
    try:
        libre_origen = '/homologacion_psico_2_6.xls'
        fuentecabecera = easyxf(
            'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
        fuentenormal = easyxf(
            'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')

        output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
        # liborigen = xlrd.open_workbook(output_folder + libre_origen)
        libdestino = xlwt.Workbook()
        hojadestino = libdestino.add_sheet('Sheet1')
        fil = 0
        columnas = [(u"CEDULA", 6000),
                    (u"APELLIDOS Y NOMBRES", 6000),
                    (u"HORAS PRACTICAS", 6000),
                    (u"HORAS VINCULACION", 6000),
                    (u"OBSERVACIÓN", 6000)
                    ]
        for col_num in range(len(columnas)):
            hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
            hojadestino.col(col_num).width = columnas[col_num][1]
        fila = 1
        titulacion = 0

        lin = 0
        # miarchivo = openpyxl.load_workbook("CEDULA_COMUNICACION_final.xlsx")
        miarchivo = openpyxl.load_workbook("psicclinea.xlsx")
        # miarchivo = openpyxl.load_workbook("prueba_2.xlsx")

        ws = miarchivo.get_sheet_by_name("sexto")
        worksheet = ws
        c = 0
        cont = 0
        periodo_id = 224
        carrera_id = 132
        mallaantigua_id = 204
        mallanueva_id = 479

        for row in worksheet.iter_rows(min_row=0):
            if lin >= 0:
                currentValues, cadena = [], ''
                for cell in row:
                    cadena += str(cell.value) + ' '
                    currentValues.append(str(cell.value))
                identificacion = currentValues[0]

                if not identificacion:
                    break

                matricula = Matricula.objects.filter(status=True, inscripcion__carrera__id=carrera_id,
                                                     inscripcion__persona__cedula=identificacion).first()
                cont += 1
                matricula.pasoayuda = True
                matricula.save()
                print(u"%s - %s" % (matricula, cont))
                inscripcion = matricula.inscripcion
                hojadestino.write(fila, 0, matricula.inscripcion.persona.identificacion(), fuentenormal)
                hojadestino.write(fila, 1, matricula.inscripcion.persona.nombre_completo(), fuentenormal)
                hojadestino.write(fila, 2, matricula.nivel.periodo.nombre, fuentenormal)

                practicaspp = haber_cumplido_horas_creditos_practicas_preprofesionales(inscripcion.id)
                horasvinculacion = haber_cumplido_horas_creditos_vinculacion(inscripcion.id)

                if InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                   malla_id=mallaantigua_id).exists():
                    imantigua = \
                        InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                        malla_id=mallaantigua_id)[0]
                    imantigua.status = False
                    imantigua.save()
                    print(u"Desactiva antigua inscripcion -----------------------------")

                if not InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                       malla_id=mallanueva_id).exists():
                    imnueva = InscripcionMalla(inscripcion=inscripcion, malla_id=mallanueva_id)
                    imnueva.save()
                    print(u"Crea nueva inscripcion -----------------------------")

                equivalencias = TablaEquivalenciaAsignaturas.objects.filter(status=True,
                                                                            asignaturamalla__malla_id=mallaantigua_id).order_by(
                    'asignaturamallasalto__nivelmalla__orden')

                cont_asig_vinculacion_aprobadas = 0
                horasfalta = 0
                fechainicioitinerario = None
                fechafinitinerario = None

                temporal = []
                for e in equivalencias:
                    temporal.append(
                        [inscripcion.recordacademico_set.filter(status=True,
                                                                asignaturamalla=e.asignaturamalla).first(),
                         e])

                for t in temporal:
                    recordantiguo, equivalencia = t

                    old, new = equivalencia.asignaturamalla, equivalencia.asignaturamallasalto
                    print(f"Nueva - {equivalencia.asignaturamallasalto}")

                    if recordantiguo:
                        print(f"Anterior - {equivalencia.asignaturamalla}")
                        print(f"Record antiguo: {recordantiguo}")

                        recordantiguo.status = False
                        recordantiguo.save(update_asignaturamalla=False)

                        recordnuevo, homologada = None, False
                        if equivalencia.asignaturamallasalto_id in (10646, 10649, 10654, 10639):
                            observaciones = recordantiguo.observaciones + " Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15"
                            homologada = True
                        else:
                            observaciones = recordantiguo.observaciones + " Migración con base a RESOLUCIÓN CGA-SO-9-2023-NO15 Y RESOLUCIÓN OCS-SO-18-2023-NO14"

                        if recordnuevo := inscripcion.recordacademico_set.filter(asignaturamalla=new,
                                                                                 status=True).first():
                            recordnuevo.matriculas = recordantiguo.matriculas
                            recordnuevo.asignaturamalla = equivalencia.asignaturamallasalto
                            recordnuevo.asignatura = equivalencia.asignaturamallasalto.asignatura
                            recordnuevo.asignaturaold = recordantiguo.asignaturamalla.asignatura
                            recordnuevo.nota = recordantiguo.nota
                            recordnuevo.asistencia = recordantiguo.asistencia
                            recordnuevo.sinasistencia = recordantiguo.sinasistencia
                            recordnuevo.fecha = recordantiguo.fecha
                            recordnuevo.noaplica = recordantiguo.noaplica
                            recordnuevo.aprobada = recordantiguo.aprobada
                            recordnuevo.convalidacion = recordantiguo.convalidacion
                            recordnuevo.pendiente = recordantiguo.pendiente
                            recordnuevo.creditos = equivalencia.asignaturamallasalto.creditos
                            recordnuevo.horas = equivalencia.asignaturamallasalto.horas
                            recordnuevo.valida = recordantiguo.valida
                            recordnuevo.validapromedio = recordantiguo.validapromedio
                            recordnuevo.observaciones = observaciones
                            recordnuevo.homologada = homologada
                            recordnuevo.materiaregular = recordantiguo.materiaregular
                            recordnuevo.materiacurso = None
                            recordnuevo.completonota = recordantiguo.completonota
                            recordnuevo.completoasistencia = recordantiguo.completoasistencia
                            recordnuevo.fechainicio = recordantiguo.fechainicio
                            recordnuevo.fechafin = recordantiguo.fechafin
                            recordnuevo.suficiencia = recordantiguo.suficiencia
                            recordnuevo.asignaturamallahistorico = recordantiguo.asignaturamalla
                            recordnuevo.reverso = False
                            recordnuevo.save()
                        else:
                            recordnuevo = RecordAcademico(inscripcion=inscripcion,
                                                          matriculas=recordantiguo.matriculas,
                                                          asignaturamalla=equivalencia.asignaturamallasalto,
                                                          asignatura=equivalencia.asignaturamallasalto.asignatura,
                                                          asignaturaold_id=recordantiguo.asignatura.id,
                                                          nota=recordantiguo.nota,
                                                          asistencia=recordantiguo.asistencia,
                                                          sinasistencia=recordantiguo.sinasistencia,
                                                          fecha=recordantiguo.fecha,
                                                          noaplica=recordantiguo.noaplica,
                                                          aprobada=recordantiguo.aprobada,
                                                          convalidacion=recordantiguo.convalidacion,
                                                          pendiente=recordantiguo.pendiente,
                                                          creditos=equivalencia.asignaturamallasalto.creditos,
                                                          horas=equivalencia.asignaturamallasalto.horas,
                                                          valida=recordantiguo.valida,
                                                          validapromedio=recordantiguo.validapromedio,
                                                          observaciones=observaciones,
                                                          homologada=homologada,
                                                          materiaregular=recordantiguo.materiaregular,
                                                          materiacurso=None,
                                                          completonota=recordantiguo.completonota,
                                                          completoasistencia=recordantiguo.completoasistencia,
                                                          fechainicio=recordantiguo.fechainicio,
                                                          fechafin=recordantiguo.fechafin,
                                                          suficiencia=recordantiguo.suficiencia,
                                                          asignaturamallahistorico_id=recordantiguo.asignaturamalla.id,
                                                          reverso=False)
                            recordnuevo.save()
                            print(u"Crea nuevo record %s" % recordnuevo)

                        if recordnuevo:
                            historicos = HistoricoRecordAcademico.objects.filter(status=True,
                                                                                 recordacademico=recordantiguo).update(
                                recordacademico=recordnuevo, creditos=recordnuevo.creditos, horas=recordnuevo.horas,
                                homologada=recordnuevo.homologada)
                            respaldo = RespaldoRecordAcademico.objects.filter(status=True,
                                                                              recordacademicooriginal=recordantiguo)

                            if not respaldo.exists():
                                respaldorecord = RespaldoRecordAcademico(recordacademicooriginal=recordantiguo,
                                                                         recordacademiconuevo=recordnuevo)
                                respaldorecord.save()
                            else:
                                respaldorecord = respaldo[0]
                                respaldorecord.recordacademiconuevo = recordnuevo
                                respaldorecord.save()

                            if equivalencia.asignaturamallasalto_id in [10646, 10649, 10654]:
                                if not practicaspp:
                                    if recordnuevo.aprobada:
                                        profesor = None
                                        if recordnuevo.materiaregular:
                                            profesor = recordnuevo.materiaregular.profesor_principal()
                                        elif recordnuevo.materiacurso:
                                            profesor = recordnuevo.materiaregular.profesor()
                                        if equivalencia.asignaturamallasalto_id == 10646:
                                            itinerariosexto = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=6)

                                            isextonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                       malla_id=mallanueva_id,
                                                                                       nivel_id=6)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(
                                                status=True,
                                                inscripcion=inscripcion,
                                                estadosolicitud__in=[
                                                    1, 2, 4,
                                                    5, 6],
                                                itinerariomalla=itinerariosexto).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerariosexto).exists()
                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        actividad__itinerariomalla=itinerariosexto).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                        fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                        numerohora=isextonuevo.horas_practicas,
                                                        nivelmalla=isextonuevo.nivel,
                                                        tiposolicitud=1,
                                                        estadosolicitud=2,
                                                        tipo=1,
                                                        itinerariomalla=isextonuevo,
                                                        supervisor=profesor,
                                                        tutorunemi=profesor,
                                                        fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                        tipoinstitucion=1,
                                                        sectoreconomico=6,
                                                        empresaempleadora_id=3,
                                                        culminada=True,
                                                        fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                        lugarpractica_id=2,
                                                        observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                    )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(
                                                status=True,
                                                itinerario=itinerariosexto).update(
                                                itinerario=isextonuevo)

                                        if equivalencia.asignaturamallasalto_id == 10649:
                                            itinerarioseptimo = ItinerariosMalla.objects.get(status=True,
                                                                                             malla_id=mallaantigua_id,
                                                                                             nivel_id=7)
                                            iseptimonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                         malla_id=mallanueva_id,
                                                                                         nivel_id=7)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(
                                                status=True,
                                                inscripcion=inscripcion,
                                                estadosolicitud__in=[
                                                    1, 2, 4,
                                                    5, 6],
                                                itinerariomalla=itinerarioseptimo).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerarioseptimo).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        actividad__itinerariomalla=itinerarioseptimo).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                        fechahasta=recordnuevo.materiaregular.nivel.periodo.fin,
                                                        numerohora=iseptimonuevo.horas_practicas,
                                                        nivelmalla=iseptimonuevo.nivel,
                                                        tiposolicitud=1,
                                                        estadosolicitud=2,
                                                        tipo=1,
                                                        itinerariomalla=iseptimonuevo,
                                                        supervisor=profesor,
                                                        tutorunemi=profesor,
                                                        fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                        tipoinstitucion=1,
                                                        sectoreconomico=6,
                                                        empresaempleadora_id=3,
                                                        culminada=True,
                                                        fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                        lugarpractica_id=2,
                                                        observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                    )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(
                                                status=True,
                                                itinerario=itinerarioseptimo).update(
                                                itinerario=iseptimonuevo)

                                        if equivalencia.asignaturamallasalto_id == 10654:
                                            itinerariooctavo = ItinerariosMalla.objects.get(status=True,
                                                                                            malla_id=mallaantigua_id,
                                                                                            nivel_id=8)
                                            ioctavonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=8)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(
                                                status=True,
                                                inscripcion=inscripcion,
                                                estadosolicitud__in=[
                                                    1, 2, 4,
                                                    5, 6],
                                                itinerariomalla=itinerariooctavo).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerariooctavo).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        actividad__itinerariomalla=itinerariooctavo).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                        fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                        numerohora=ioctavonuevo.horas_practicas,
                                                        nivelmalla=ioctavonuevo.nivel,
                                                        tiposolicitud=1,
                                                        estadosolicitud=2,
                                                        tipo=1,
                                                        itinerariomalla=ioctavonuevo,
                                                        supervisor=profesor,
                                                        tutorunemi=profesor,
                                                        fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                        tipoinstitucion=1,
                                                        sectoreconomico=6,
                                                        empresaempleadora_id=3,
                                                        culminada=True,
                                                        fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                        lugarpractica_id=2,
                                                        observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                    )
                                                    nuevapractica.save()
                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(
                                                status=True,
                                                itinerario=itinerariooctavo).update(
                                                itinerario=ioctavonuevo)

                            if equivalencia.asignaturamallasalto_id in [10639, 10654]:
                                if not horasvinculacion:
                                    if recordnuevo.aprobada:
                                        if equivalencia.asignaturamallasalto_id == 10639:
                                            cont_asig_vinculacion_aprobadas += 1

                                        if equivalencia.asignaturamallasalto_id == 10654:
                                            cont_asig_vinculacion_aprobadas += 1
                                        fechainicioitinerario = recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date()
                                        fechafinitinerario = recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date()

                            print(u"Record actualizado %s" % recordnuevo)
                    else:
                        hojadestino.write(fila, 5, "NO ENCONTRO RECORD ANTIGUO %s" % equivalencia.asignaturamalla)
                        fila += 1

                if cont_asig_vinculacion_aprobadas != 0:
                    if cont_asig_vinculacion_aprobadas == 1:
                        horasfalta = 144
                    if cont_asig_vinculacion_aprobadas == 2:
                        horasfalta = 160
                    horasfalta = horasfalta - inscripcion.numero_horas_proyectos_vinculacion()
                    vinculacion = ParticipantesMatrices(status=True,
                                                        matrizevidencia_id=2,
                                                        proyecto_id=601,
                                                        inscripcion=inscripcion,
                                                        horas=horasfalta,
                                                        registrohorasdesde=fechainicioitinerario,
                                                        registrohorashasta=fechafinitinerario,
                                                        estado=1
                                                        )
                    vinculacion.save()

                practicasppf = inscripcion.numero_horas_practicas_pre_profesionales()
                hojadestino.write(fila, 3, practicasppf, fuentenormal)
                horasvinculacionf = inscripcion.numero_horas_proyectos_vinculacion()
                hojadestino.write(fila, 4, horasvinculacionf, fuentenormal)

                fila += 1

                time.sleep(3)

            lin += 1

        libdestino.save(output_folder + libre_origen)
        print(output_folder + libre_origen)
        print("Proceso finalizado. . .")

    except Exception as ex:
        noti = Notificacion(titulo='Error',
                            cuerpo='Ha ocurrido un error {} - Error en la linea {}'.format(
                                ex, sys.exc_info()[-1].tb_lineno),
                            destinatario_id=29898, url="",
                            prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                            tipo=2, en_proceso=False, error=True)
        noti.save()
        transaction.set_rollback(True)
        print('error: %s' % ex)
        hojadestino.write(fila, 3, str(ex))
        fila += 1

def homologacion_psicologia7():
    try:
        libre_origen = '/homologacion_psico_2_7.xls'
        fuentecabecera = easyxf(
            'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
        fuentenormal = easyxf(
            'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')

        output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
        # liborigen = xlrd.open_workbook(output_folder + libre_origen)
        libdestino = xlwt.Workbook()
        hojadestino = libdestino.add_sheet('Sheet1')
        fil = 0
        columnas = [(u"CEDULA", 6000),
                    (u"APELLIDOS Y NOMBRES", 6000),
                    (u"HORAS PRACTICAS", 6000),
                    (u"HORAS VINCULACION", 6000),
                    (u"OBSERVACIÓN", 6000)
                    ]
        for col_num in range(len(columnas)):
            hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
            hojadestino.col(col_num).width = columnas[col_num][1]
        fila = 1
        titulacion = 0

        lin = 0
        # miarchivo = openpyxl.load_workbook("CEDULA_COMUNICACION_final.xlsx")
        miarchivo = openpyxl.load_workbook("psicclinea.xlsx")
        # miarchivo = openpyxl.load_workbook("prueba_2.xlsx")

        ws = miarchivo.get_sheet_by_name("septimo")
        worksheet = ws
        c = 0
        cont = 0
        periodo_id = 224
        carrera_id = 132
        mallaantigua_id = 204
        mallanueva_id = 479

        for row in worksheet.iter_rows(min_row=0):
            if lin >= 0:
                currentValues, cadena = [], ''
                for cell in row:
                    cadena += str(cell.value) + ' '
                    currentValues.append(str(cell.value))
                identificacion = currentValues[0]

                if not identificacion:
                    break

                matricula = Matricula.objects.filter(status=True, inscripcion__carrera__id=carrera_id,
                                                     inscripcion__persona__cedula=identificacion).first()
                cont += 1
                matricula.pasoayuda = True
                matricula.save()
                print(u"%s - %s" % (matricula, cont))
                inscripcion = matricula.inscripcion
                hojadestino.write(fila, 0, matricula.inscripcion.persona.identificacion(), fuentenormal)
                hojadestino.write(fila, 1, matricula.inscripcion.persona.nombre_completo(), fuentenormal)
                hojadestino.write(fila, 2, matricula.nivel.periodo.nombre, fuentenormal)

                practicaspp = haber_cumplido_horas_creditos_practicas_preprofesionales(inscripcion.id)
                horasvinculacion = haber_cumplido_horas_creditos_vinculacion(inscripcion.id)

                if InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                   malla_id=mallaantigua_id).exists():
                    imantigua = \
                        InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                        malla_id=mallaantigua_id)[0]
                    imantigua.status = False
                    imantigua.save()
                    print(u"Desactiva antigua inscripcion -----------------------------")

                if not InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                       malla_id=mallanueva_id).exists():
                    imnueva = InscripcionMalla(inscripcion=inscripcion, malla_id=mallanueva_id)
                    imnueva.save()
                    print(u"Crea nueva inscripcion -----------------------------")

                equivalencias = TablaEquivalenciaAsignaturas.objects.filter(status=True,
                                                                            asignaturamalla__malla_id=mallaantigua_id).order_by(
                    'asignaturamallasalto__nivelmalla__orden')

                cont_asig_vinculacion_aprobadas = 0
                horasfalta = 0
                fechainicioitinerario = None
                fechafinitinerario = None

                temporal = []
                for e in equivalencias:
                    temporal.append(
                        [inscripcion.recordacademico_set.filter(status=True,
                                                                asignaturamalla=e.asignaturamalla).first(),
                         e])

                for t in temporal:
                    recordantiguo, equivalencia = t

                    old, new = equivalencia.asignaturamalla, equivalencia.asignaturamallasalto
                    print(f"Nueva - {equivalencia.asignaturamallasalto}")

                    if recordantiguo:
                        print(f"Anterior - {equivalencia.asignaturamalla}")
                        print(f"Record antiguo: {recordantiguo}")

                        recordantiguo.status = False
                        recordantiguo.save(update_asignaturamalla=False)

                        recordnuevo, homologada = None, False
                        if equivalencia.asignaturamallasalto_id in (10646, 10649, 10654, 10639):
                            observaciones = recordantiguo.observaciones + " Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15"
                            homologada = True
                        else:
                            observaciones = recordantiguo.observaciones + " Migración con base a RESOLUCIÓN CGA-SO-9-2023-NO15 Y RESOLUCIÓN OCS-SO-18-2023-NO14"

                        if recordnuevo := inscripcion.recordacademico_set.filter(asignaturamalla=new,
                                                                                 status=True).first():
                            recordnuevo.matriculas = recordantiguo.matriculas
                            recordnuevo.asignaturamalla = equivalencia.asignaturamallasalto
                            recordnuevo.asignatura = equivalencia.asignaturamallasalto.asignatura
                            recordnuevo.asignaturaold = recordantiguo.asignaturamalla.asignatura
                            recordnuevo.nota = recordantiguo.nota
                            recordnuevo.asistencia = recordantiguo.asistencia
                            recordnuevo.sinasistencia = recordantiguo.sinasistencia
                            recordnuevo.fecha = recordantiguo.fecha
                            recordnuevo.noaplica = recordantiguo.noaplica
                            recordnuevo.aprobada = recordantiguo.aprobada
                            recordnuevo.convalidacion = recordantiguo.convalidacion
                            recordnuevo.pendiente = recordantiguo.pendiente
                            recordnuevo.creditos = equivalencia.asignaturamallasalto.creditos
                            recordnuevo.horas = equivalencia.asignaturamallasalto.horas
                            recordnuevo.valida = recordantiguo.valida
                            recordnuevo.validapromedio = recordantiguo.validapromedio
                            recordnuevo.observaciones = observaciones
                            recordnuevo.homologada = homologada
                            recordnuevo.materiaregular = recordantiguo.materiaregular
                            recordnuevo.materiacurso = None
                            recordnuevo.completonota = recordantiguo.completonota
                            recordnuevo.completoasistencia = recordantiguo.completoasistencia
                            recordnuevo.fechainicio = recordantiguo.fechainicio
                            recordnuevo.fechafin = recordantiguo.fechafin
                            recordnuevo.suficiencia = recordantiguo.suficiencia
                            recordnuevo.asignaturamallahistorico = recordantiguo.asignaturamalla
                            recordnuevo.reverso = False
                            recordnuevo.save()
                        else:
                            recordnuevo = RecordAcademico(inscripcion=inscripcion,
                                                          matriculas=recordantiguo.matriculas,
                                                          asignaturamalla=equivalencia.asignaturamallasalto,
                                                          asignatura=equivalencia.asignaturamallasalto.asignatura,
                                                          asignaturaold_id=recordantiguo.asignatura.id,
                                                          nota=recordantiguo.nota,
                                                          asistencia=recordantiguo.asistencia,
                                                          sinasistencia=recordantiguo.sinasistencia,
                                                          fecha=recordantiguo.fecha,
                                                          noaplica=recordantiguo.noaplica,
                                                          aprobada=recordantiguo.aprobada,
                                                          convalidacion=recordantiguo.convalidacion,
                                                          pendiente=recordantiguo.pendiente,
                                                          creditos=equivalencia.asignaturamallasalto.creditos,
                                                          horas=equivalencia.asignaturamallasalto.horas,
                                                          valida=recordantiguo.valida,
                                                          validapromedio=recordantiguo.validapromedio,
                                                          observaciones=observaciones,
                                                          homologada=homologada,
                                                          materiaregular=recordantiguo.materiaregular,
                                                          materiacurso=None,
                                                          completonota=recordantiguo.completonota,
                                                          completoasistencia=recordantiguo.completoasistencia,
                                                          fechainicio=recordantiguo.fechainicio,
                                                          fechafin=recordantiguo.fechafin,
                                                          suficiencia=recordantiguo.suficiencia,
                                                          asignaturamallahistorico_id=recordantiguo.asignaturamalla.id,
                                                          reverso=False)
                            recordnuevo.save()
                            print(u"Crea nuevo record %s" % recordnuevo)

                        if recordnuevo:
                            historicos = HistoricoRecordAcademico.objects.filter(status=True,
                                                                                 recordacademico=recordantiguo).update(
                                recordacademico=recordnuevo, creditos=recordnuevo.creditos, horas=recordnuevo.horas,
                                homologada=recordnuevo.homologada)
                            respaldo = RespaldoRecordAcademico.objects.filter(status=True,
                                                                              recordacademicooriginal=recordantiguo)

                            if not respaldo.exists():
                                respaldorecord = RespaldoRecordAcademico(recordacademicooriginal=recordantiguo,
                                                                         recordacademiconuevo=recordnuevo)
                                respaldorecord.save()
                            else:
                                respaldorecord = respaldo[0]
                                respaldorecord.recordacademiconuevo = recordnuevo
                                respaldorecord.save()

                            if equivalencia.asignaturamallasalto_id in [10646, 10649, 10654]:
                                if not practicaspp:
                                    if recordnuevo.aprobada:
                                        profesor = None
                                        if recordnuevo.materiaregular:
                                            profesor = recordnuevo.materiaregular.profesor_principal()
                                        elif recordnuevo.materiacurso:
                                            profesor = recordnuevo.materiaregular.profesor()
                                        if equivalencia.asignaturamallasalto_id == 10646:
                                            itinerariosexto = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=6)

                                            isextonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                       malla_id=mallanueva_id,
                                                                                       nivel_id=6)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(
                                                status=True,
                                                inscripcion=inscripcion,
                                                estadosolicitud__in=[
                                                    1, 2, 4,
                                                    5, 6],
                                                itinerariomalla=itinerariosexto).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerariosexto).exists()
                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        actividad__itinerariomalla=itinerariosexto).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                        fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                        numerohora=isextonuevo.horas_practicas,
                                                        nivelmalla=isextonuevo.nivel,
                                                        tiposolicitud=1,
                                                        estadosolicitud=2,
                                                        tipo=1,
                                                        itinerariomalla=isextonuevo,
                                                        supervisor=profesor,
                                                        tutorunemi=profesor,
                                                        fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                        tipoinstitucion=1,
                                                        sectoreconomico=6,
                                                        empresaempleadora_id=3,
                                                        culminada=True,
                                                        fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                        lugarpractica_id=2,
                                                        observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                    )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(
                                                status=True,
                                                itinerario=itinerariosexto).update(
                                                itinerario=isextonuevo)

                                        if equivalencia.asignaturamallasalto_id == 10649:
                                            itinerarioseptimo = ItinerariosMalla.objects.get(status=True,
                                                                                             malla_id=mallaantigua_id,
                                                                                             nivel_id=7)
                                            iseptimonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                         malla_id=mallanueva_id,
                                                                                         nivel_id=7)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(
                                                status=True,
                                                inscripcion=inscripcion,
                                                estadosolicitud__in=[
                                                    1, 2, 4,
                                                    5, 6],
                                                itinerariomalla=itinerarioseptimo).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerarioseptimo).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        actividad__itinerariomalla=itinerarioseptimo).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                        fechahasta=recordnuevo.materiaregular.nivel.periodo.fin,
                                                        numerohora=iseptimonuevo.horas_practicas,
                                                        nivelmalla=iseptimonuevo.nivel,
                                                        tiposolicitud=1,
                                                        estadosolicitud=2,
                                                        tipo=1,
                                                        itinerariomalla=iseptimonuevo,
                                                        supervisor=profesor,
                                                        tutorunemi=profesor,
                                                        fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                        tipoinstitucion=1,
                                                        sectoreconomico=6,
                                                        empresaempleadora_id=3,
                                                        culminada=True,
                                                        fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                        lugarpractica_id=2,
                                                        observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                    )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(
                                                status=True,
                                                itinerario=itinerarioseptimo).update(
                                                itinerario=iseptimonuevo)

                                        if equivalencia.asignaturamallasalto_id == 10654:
                                            itinerariooctavo = ItinerariosMalla.objects.get(status=True,
                                                                                            malla_id=mallaantigua_id,
                                                                                            nivel_id=8)
                                            ioctavonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=8)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(
                                                status=True,
                                                inscripcion=inscripcion,
                                                estadosolicitud__in=[
                                                    1, 2, 4,
                                                    5, 6],
                                                itinerariomalla=itinerariooctavo).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerariooctavo).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        actividad__itinerariomalla=itinerariooctavo).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                        fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                        numerohora=ioctavonuevo.horas_practicas,
                                                        nivelmalla=ioctavonuevo.nivel,
                                                        tiposolicitud=1,
                                                        estadosolicitud=2,
                                                        tipo=1,
                                                        itinerariomalla=ioctavonuevo,
                                                        supervisor=profesor,
                                                        tutorunemi=profesor,
                                                        fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                        tipoinstitucion=1,
                                                        sectoreconomico=6,
                                                        empresaempleadora_id=3,
                                                        culminada=True,
                                                        fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                        lugarpractica_id=2,
                                                        observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                    )
                                                    nuevapractica.save()
                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(
                                                status=True,
                                                itinerario=itinerariooctavo).update(
                                                itinerario=ioctavonuevo)

                            if equivalencia.asignaturamallasalto_id in [10639, 10654]:
                                if not horasvinculacion:
                                    if recordnuevo.aprobada:
                                        if equivalencia.asignaturamallasalto_id == 10639:
                                            cont_asig_vinculacion_aprobadas += 1

                                        if equivalencia.asignaturamallasalto_id == 10654:
                                            cont_asig_vinculacion_aprobadas += 1
                                        fechainicioitinerario = recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date()
                                        fechafinitinerario = recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date()

                            print(u"Record actualizado %s" % recordnuevo)
                    else:
                        hojadestino.write(fila, 5, "NO ENCONTRO RECORD ANTIGUO %s" % equivalencia.asignaturamalla)
                        fila += 1

                if cont_asig_vinculacion_aprobadas != 0:
                    if cont_asig_vinculacion_aprobadas == 1:
                        horasfalta = 144
                    if cont_asig_vinculacion_aprobadas == 2:
                        horasfalta = 160
                    horasfalta = horasfalta - inscripcion.numero_horas_proyectos_vinculacion()
                    vinculacion = ParticipantesMatrices(status=True,
                                                        matrizevidencia_id=2,
                                                        proyecto_id=601,
                                                        inscripcion=inscripcion,
                                                        horas=horasfalta,
                                                        registrohorasdesde=fechainicioitinerario,
                                                        registrohorashasta=fechafinitinerario,
                                                        estado=1
                                                        )
                    vinculacion.save()

                practicasppf = inscripcion.numero_horas_practicas_pre_profesionales()
                hojadestino.write(fila, 3, practicasppf, fuentenormal)
                horasvinculacionf = inscripcion.numero_horas_proyectos_vinculacion()
                hojadestino.write(fila, 4, horasvinculacionf, fuentenormal)

                fila += 1

                time.sleep(3)

            lin += 1

        libdestino.save(output_folder + libre_origen)
        print(output_folder + libre_origen)
        print("Proceso finalizado. . .")

    except Exception as ex:
        noti = Notificacion(titulo='Error',
                            cuerpo='Ha ocurrido un error {} - Error en la linea {}'.format(
                                ex, sys.exc_info()[-1].tb_lineno),
                            destinatario_id=29898, url="",
                            prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                            tipo=2, en_proceso=False, error=True)
        noti.save()
        transaction.set_rollback(True)
        print('error: %s' % ex)
        hojadestino.write(fila, 3, str(ex))
        fila += 1

def homologacion_psicologia8():
    try:
        libre_origen = '/homologacion_psico_2_8.xls'
        fuentecabecera = easyxf(
            'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
        fuentenormal = easyxf(
            'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')

        output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
        # liborigen = xlrd.open_workbook(output_folder + libre_origen)
        libdestino = xlwt.Workbook()
        hojadestino = libdestino.add_sheet('Sheet1')
        fil = 0
        columnas = [(u"CEDULA", 6000),
                    (u"APELLIDOS Y NOMBRES", 6000),
                    (u"HORAS PRACTICAS", 6000),
                    (u"HORAS VINCULACION", 6000),
                    (u"OBSERVACIÓN", 6000)
                    ]
        for col_num in range(len(columnas)):
            hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
            hojadestino.col(col_num).width = columnas[col_num][1]
        fila = 1
        titulacion = 0

        lin = 0
        # miarchivo = openpyxl.load_workbook("CEDULA_COMUNICACION_final.xlsx")
        miarchivo = openpyxl.load_workbook("psiclinea.xlsx")
        # miarchivo = openpyxl.load_workbook("prueba_2.xlsx")

        ws = miarchivo.get_sheet_by_name("octavo")
        worksheet = ws
        c = 0
        cont = 0
        periodo_id = 224
        carrera_id = 132
        mallaantigua_id = 204
        mallanueva_id = 479

        for row in worksheet.iter_rows(min_row=0):
            if lin >= 0:
                currentValues, cadena = [], ''
                for cell in row:
                    cadena += str(cell.value) + ' '
                    currentValues.append(str(cell.value))
                identificacion = currentValues[0]

                if not identificacion:
                    break

                matricula = Matricula.objects.filter(status=True, inscripcion__carrera__id=carrera_id,
                                                     inscripcion__persona__cedula=identificacion).first()
                cont += 1
                matricula.pasoayuda = True
                matricula.save()
                print(u"%s - %s" % (matricula, cont))
                inscripcion = matricula.inscripcion
                hojadestino.write(fila, 0, matricula.inscripcion.persona.identificacion(), fuentenormal)
                hojadestino.write(fila, 1, matricula.inscripcion.persona.nombre_completo(), fuentenormal)
                hojadestino.write(fila, 2, matricula.nivel.periodo.nombre, fuentenormal)

                practicaspp = haber_cumplido_horas_creditos_practicas_preprofesionales(inscripcion.id)
                horasvinculacion = haber_cumplido_horas_creditos_vinculacion(inscripcion.id)

                if InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                   malla_id=mallaantigua_id).exists():
                    imantigua = \
                        InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                        malla_id=mallaantigua_id)[0]
                    imantigua.status = False
                    imantigua.save()
                    print(u"Desactiva antigua inscripcion -----------------------------")

                if not InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                       malla_id=mallanueva_id).exists():
                    imnueva = InscripcionMalla(inscripcion=inscripcion, malla_id=mallanueva_id)
                    imnueva.save()
                    print(u"Crea nueva inscripcion -----------------------------")

                equivalencias = TablaEquivalenciaAsignaturas.objects.filter(status=True,
                                                                            asignaturamalla__malla_id=mallaantigua_id).order_by(
                    'asignaturamallasalto__nivelmalla__orden')

                cont_asig_vinculacion_aprobadas = 0
                horasfalta = 0
                fechainicioitinerario = None
                fechafinitinerario = None

                temporal = []
                for e in equivalencias:
                    temporal.append(
                        [inscripcion.recordacademico_set.filter(status=True,
                                                                asignaturamalla=e.asignaturamalla).first(),
                         e])

                for t in temporal:
                    recordantiguo, equivalencia = t

                    old, new = equivalencia.asignaturamalla, equivalencia.asignaturamallasalto
                    print(f"Nueva - {equivalencia.asignaturamallasalto}")

                    if recordantiguo:
                        print(f"Anterior - {equivalencia.asignaturamalla}")
                        print(f"Record antiguo: {recordantiguo}")

                        recordantiguo.status = False
                        recordantiguo.save(update_asignaturamalla=False)

                        recordnuevo, homologada = None, False
                        if equivalencia.asignaturamallasalto_id in (10646, 10649, 10654, 10639):
                            observaciones = recordantiguo.observaciones + " Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15"
                            homologada = True
                        else:
                            observaciones = recordantiguo.observaciones + " Migración con base a RESOLUCIÓN CGA-SO-9-2023-NO15 Y RESOLUCIÓN OCS-SO-18-2023-NO14"

                        if recordnuevo := inscripcion.recordacademico_set.filter(asignaturamalla=new,
                                                                                 status=True).first():
                            recordnuevo.matriculas = recordantiguo.matriculas
                            recordnuevo.asignaturamalla = equivalencia.asignaturamallasalto
                            recordnuevo.asignatura = equivalencia.asignaturamallasalto.asignatura
                            recordnuevo.asignaturaold = recordantiguo.asignaturamalla.asignatura
                            recordnuevo.nota = recordantiguo.nota
                            recordnuevo.asistencia = recordantiguo.asistencia
                            recordnuevo.sinasistencia = recordantiguo.sinasistencia
                            recordnuevo.fecha = recordantiguo.fecha
                            recordnuevo.noaplica = recordantiguo.noaplica
                            recordnuevo.aprobada = recordantiguo.aprobada
                            recordnuevo.convalidacion = recordantiguo.convalidacion
                            recordnuevo.pendiente = recordantiguo.pendiente
                            recordnuevo.creditos = equivalencia.asignaturamallasalto.creditos
                            recordnuevo.horas = equivalencia.asignaturamallasalto.horas
                            recordnuevo.valida = recordantiguo.valida
                            recordnuevo.validapromedio = recordantiguo.validapromedio
                            recordnuevo.observaciones = observaciones
                            recordnuevo.homologada = homologada
                            recordnuevo.materiaregular = recordantiguo.materiaregular
                            recordnuevo.materiacurso = None
                            recordnuevo.completonota = recordantiguo.completonota
                            recordnuevo.completoasistencia = recordantiguo.completoasistencia
                            recordnuevo.fechainicio = recordantiguo.fechainicio
                            recordnuevo.fechafin = recordantiguo.fechafin
                            recordnuevo.suficiencia = recordantiguo.suficiencia
                            recordnuevo.asignaturamallahistorico = recordantiguo.asignaturamalla
                            recordnuevo.reverso = False
                            recordnuevo.save()
                        else:
                            recordnuevo = RecordAcademico(inscripcion=inscripcion,
                                                          matriculas=recordantiguo.matriculas,
                                                          asignaturamalla=equivalencia.asignaturamallasalto,
                                                          asignatura=equivalencia.asignaturamallasalto.asignatura,
                                                          asignaturaold_id=recordantiguo.asignatura.id,
                                                          nota=recordantiguo.nota,
                                                          asistencia=recordantiguo.asistencia,
                                                          sinasistencia=recordantiguo.sinasistencia,
                                                          fecha=recordantiguo.fecha,
                                                          noaplica=recordantiguo.noaplica,
                                                          aprobada=recordantiguo.aprobada,
                                                          convalidacion=recordantiguo.convalidacion,
                                                          pendiente=recordantiguo.pendiente,
                                                          creditos=equivalencia.asignaturamallasalto.creditos,
                                                          horas=equivalencia.asignaturamallasalto.horas,
                                                          valida=recordantiguo.valida,
                                                          validapromedio=recordantiguo.validapromedio,
                                                          observaciones=observaciones,
                                                          homologada=homologada,
                                                          materiaregular=recordantiguo.materiaregular,
                                                          materiacurso=None,
                                                          completonota=recordantiguo.completonota,
                                                          completoasistencia=recordantiguo.completoasistencia,
                                                          fechainicio=recordantiguo.fechainicio,
                                                          fechafin=recordantiguo.fechafin,
                                                          suficiencia=recordantiguo.suficiencia,
                                                          asignaturamallahistorico_id=recordantiguo.asignaturamalla.id,
                                                          reverso=False)
                            recordnuevo.save()
                            print(u"Crea nuevo record %s" % recordnuevo)

                        if recordnuevo:
                            historicos = HistoricoRecordAcademico.objects.filter(status=True,
                                                                                 recordacademico=recordantiguo).update(
                                recordacademico=recordnuevo, creditos=recordnuevo.creditos, horas=recordnuevo.horas,
                                homologada=recordnuevo.homologada)
                            respaldo = RespaldoRecordAcademico.objects.filter(status=True,
                                                                              recordacademicooriginal=recordantiguo)

                            if not respaldo.exists():
                                respaldorecord = RespaldoRecordAcademico(recordacademicooriginal=recordantiguo,
                                                                         recordacademiconuevo=recordnuevo)
                                respaldorecord.save()
                            else:
                                respaldorecord = respaldo[0]
                                respaldorecord.recordacademiconuevo = recordnuevo
                                respaldorecord.save()

                            if equivalencia.asignaturamallasalto_id in [10646, 10649, 10654]:
                                if not practicaspp:
                                    if recordnuevo.aprobada:
                                        profesor = None
                                        if recordnuevo.materiaregular:
                                            profesor = recordnuevo.materiaregular.profesor_principal()
                                        elif recordnuevo.materiacurso:
                                            profesor = recordnuevo.materiaregular.profesor()
                                        if equivalencia.asignaturamallasalto_id == 10646:
                                            itinerariosexto = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallaantigua_id,
                                                                                           nivel_id=6)

                                            isextonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                       malla_id=mallanueva_id,
                                                                                       nivel_id=6)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(
                                                status=True,
                                                inscripcion=inscripcion,
                                                estadosolicitud__in=[
                                                    1, 2, 4,
                                                    5, 6],
                                                itinerariomalla=itinerariosexto).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerariosexto).exists()
                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        actividad__itinerariomalla=itinerariosexto).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                        fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                        numerohora=isextonuevo.horas_practicas,
                                                        nivelmalla=isextonuevo.nivel,
                                                        tiposolicitud=1,
                                                        estadosolicitud=2,
                                                        tipo=1,
                                                        itinerariomalla=isextonuevo,
                                                        supervisor=profesor,
                                                        tutorunemi=profesor,
                                                        fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                        tipoinstitucion=1,
                                                        sectoreconomico=6,
                                                        empresaempleadora_id=3,
                                                        culminada=True,
                                                        fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                        lugarpractica_id=2,
                                                        observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                    )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(
                                                status=True,
                                                itinerario=itinerariosexto).update(
                                                itinerario=isextonuevo)

                                        if equivalencia.asignaturamallasalto_id == 10649:
                                            itinerarioseptimo = ItinerariosMalla.objects.get(status=True,
                                                                                             malla_id=mallaantigua_id,
                                                                                             nivel_id=7)
                                            iseptimonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                         malla_id=mallanueva_id,
                                                                                         nivel_id=7)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(
                                                status=True,
                                                inscripcion=inscripcion,
                                                estadosolicitud__in=[
                                                    1, 2, 4,
                                                    5, 6],
                                                itinerariomalla=itinerarioseptimo).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerarioseptimo).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        actividad__itinerariomalla=itinerarioseptimo).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                        fechahasta=recordnuevo.materiaregular.nivel.periodo.fin,
                                                        numerohora=iseptimonuevo.horas_practicas,
                                                        nivelmalla=iseptimonuevo.nivel,
                                                        tiposolicitud=1,
                                                        estadosolicitud=2,
                                                        tipo=1,
                                                        itinerariomalla=iseptimonuevo,
                                                        supervisor=profesor,
                                                        tutorunemi=profesor,
                                                        fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                        tipoinstitucion=1,
                                                        sectoreconomico=6,
                                                        empresaempleadora_id=3,
                                                        culminada=True,
                                                        fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                        lugarpractica_id=2,
                                                        observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                    )
                                                    nuevapractica.save()

                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(
                                                status=True,
                                                itinerario=itinerarioseptimo).update(
                                                itinerario=iseptimonuevo)

                                        if equivalencia.asignaturamallasalto_id == 10654:
                                            itinerariooctavo = ItinerariosMalla.objects.get(status=True,
                                                                                            malla_id=mallaantigua_id,
                                                                                            nivel_id=8)
                                            ioctavonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                        malla_id=mallanueva_id,
                                                                                        nivel_id=8)

                                            practica = PracticasPreprofesionalesInscripcion.objects.filter(
                                                status=True,
                                                inscripcion=inscripcion,
                                                estadosolicitud__in=[
                                                    1, 2, 4,
                                                    5, 6],
                                                itinerariomalla=itinerariooctavo).exists()
                                            practicarechazada = False
                                            if not practica:
                                                practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud=3,
                                                    itinerariomalla=itinerariooctavo).exists()

                                            if not practica or practicarechazada:
                                                if not PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        actividad__itinerariomalla=itinerariooctavo).exists():
                                                    nuevapractica = PracticasPreprofesionalesInscripcion(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                        fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                        numerohora=ioctavonuevo.horas_practicas,
                                                        nivelmalla=ioctavonuevo.nivel,
                                                        tiposolicitud=1,
                                                        estadosolicitud=2,
                                                        tipo=1,
                                                        itinerariomalla=ioctavonuevo,
                                                        supervisor=profesor,
                                                        tutorunemi=profesor,
                                                        fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                        tipoinstitucion=1,
                                                        sectoreconomico=6,
                                                        empresaempleadora_id=3,
                                                        culminada=True,
                                                        fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                        lugarpractica_id=2,
                                                        observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                    )
                                                    nuevapractica.save()
                                            ItinerariosActividadDetalleDistributivoCarrera.objects.filter(
                                                status=True,
                                                itinerario=itinerariooctavo).update(
                                                itinerario=ioctavonuevo)

                            if equivalencia.asignaturamallasalto_id in [10639, 10654]:
                                if not horasvinculacion:
                                    if recordnuevo.aprobada:
                                        if equivalencia.asignaturamallasalto_id == 10639:
                                            cont_asig_vinculacion_aprobadas += 1

                                        if equivalencia.asignaturamallasalto_id == 10654:
                                            cont_asig_vinculacion_aprobadas += 1
                                        fechainicioitinerario = recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date()
                                        fechafinitinerario = recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date()

                            print(u"Record actualizado %s" % recordnuevo)
                    else:
                        hojadestino.write(fila, 5, "NO ENCONTRO RECORD ANTIGUO %s" % equivalencia.asignaturamalla)
                        fila += 1

                if cont_asig_vinculacion_aprobadas != 0:
                    if cont_asig_vinculacion_aprobadas == 1:
                        horasfalta = 144
                    if cont_asig_vinculacion_aprobadas == 2:
                        horasfalta = 160
                    horasfalta = horasfalta - inscripcion.numero_horas_proyectos_vinculacion()
                    vinculacion = ParticipantesMatrices(status=True,
                                                        matrizevidencia_id=2,
                                                        proyecto_id=601,
                                                        inscripcion=inscripcion,
                                                        horas=horasfalta,
                                                        registrohorasdesde=fechainicioitinerario,
                                                        registrohorashasta=fechafinitinerario,
                                                        estado=1
                                                        )
                    vinculacion.save()

                practicasppf = inscripcion.numero_horas_practicas_pre_profesionales()
                hojadestino.write(fila, 3, practicasppf, fuentenormal)
                horasvinculacionf = inscripcion.numero_horas_proyectos_vinculacion()
                hojadestino.write(fila, 4, horasvinculacionf, fuentenormal)

                fila += 1

                time.sleep(3)

            lin += 1

        libdestino.save(output_folder + libre_origen)
        print(output_folder + libre_origen)
        print("Proceso finalizado. . .")

    except Exception as ex:
        noti = Notificacion(titulo='Error',
                            cuerpo='Ha ocurrido un error {} - Error en la linea {}'.format(
                                ex, sys.exc_info()[-1].tb_lineno),
                            destinatario_id=29898, url="",
                            prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                            tipo=2, en_proceso=False, error=True)
        noti.save()
        transaction.set_rollback(True)
        print('error: %s' % ex)
        hojadestino.write(fila, 3, str(ex))
        fila += 1

def homologacion_basica2():
    # verificar que todos los estudiantes tengan la misma malla
    cadena = ''
    linea, excluidos, conexito = 0, 0, 0
    try:
        libre_origen = '/homologacion_basica_2_2.xls'
        fuentecabecera = easyxf(
            'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
        fuentenormal = easyxf(
            'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')

        output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
        # liborigen = xlrd.open_workbook(output_folder + libre_origen)
        libdestino = xlwt.Workbook()
        hojadestino = libdestino.add_sheet('Sheet1')
        fil = 0
        columnas = [(u"CEDULA", 6000),
                    (u"APELLIDOS Y NOMBRES", 6000),
                    (u"OBSERVACIÓN", 6000),
                    (u"HORAS PRACTICAS", 6000),
                    (u"HORAS VINCULACION", 6000),
                    (u"OBSERVACIÓN", 6000)
                    ]
        for col_num in range(len(columnas)):
            hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
            hojadestino.col(col_num).width = columnas[col_num][1]
        fila = 1
        titulacion = 0

        lin = 0
        # miarchivo = openpyxl.load_workbook("CEDULA_COMUNICACION_final.xlsx")
        miarchivo = openpyxl.load_workbook("Libro1.xlsx")
        # miarchivo = openpyxl.load_workbook("prueba_2.xlsx")

        ws = miarchivo.get_sheet_by_name("Hoja1")

        worksheet = ws
        c = 0
        cont = 0
        periodo_id = 224
        carrera_id = 135
        mallaantigua_id = 208
        mallanueva_id = 490

        for row in worksheet.iter_rows(min_row=0):
            if lin >= 0:
                currentValues, cadena = [], ''
                for cell in row:
                    cadena += str(cell.value) + ' '
                    currentValues.append(str(cell.value))
                identificacion = currentValues[0]

                if not identificacion:
                    break

                with transaction.atomic():
                    matricula = Matricula.objects.filter(status=True, inscripcion__carrera__id=carrera_id,
                                                         inscripcion__persona__cedula=identificacion).first()
                    cont += 1
                    matricula.pasoayuda = True
                    matricula.save()
                    print(u"%s - %s" % (matricula, cont))
                    inscripcion = matricula.inscripcion
                    hojadestino.write(fila, 0, matricula.inscripcion.persona.identificacion(), fuentenormal)
                    hojadestino.write(fila, 1, matricula.inscripcion.persona.nombre_completo(), fuentenormal)
                    hojadestino.write(fila, 2, matricula.nivel.periodo.nombre, fuentenormal)

                    practicaspp = haber_cumplido_horas_creditos_practicas_preprofesionales(inscripcion.id)
                    horasvinculacion = haber_cumplido_horas_creditos_vinculacion(inscripcion.id)

                    if InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                       malla_id=mallaantigua_id).exists():
                        imantigua = \
                            InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                            malla_id=mallaantigua_id)[
                                0]
                        imantigua.status = False
                        imantigua.save()
                        print(u"Desactiva antigua inscripcion -----------------------------")

                    if not InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                           malla_id=mallanueva_id).exists():
                        imnueva = InscripcionMalla(inscripcion=inscripcion, malla_id=mallanueva_id)
                        imnueva.save()
                        print(u"Crea nueva inscripcion -----------------------------")

                    equivalencias = TablaEquivalenciaAsignaturas.objects.filter(status=True,
                                                                                asignaturamalla__malla_id=mallaantigua_id).order_by(
                        'asignaturamallasalto__nivelmalla__orden')
                    cont_asig_vinculacion_aprobadas = 0
                    horasfalta = 0
                    fechainicioitinerario = None
                    fechafinitinerario = None
                    for equivalencia in equivalencias:
                        print(u"nueva - %s" % equivalencia.asignaturamallasalto)
                        recordantiguo = inscripcion.recordacademico_set.filter(status=True,
                                                                               asignaturamalla_id=equivalencia.asignaturamalla_id).first()

                        if recordantiguo:
                            print(u"anterior - %s" % equivalencia.asignaturamalla)
                            print(u"Record antiguo: %s" % recordantiguo)
                            recordnuevo = None
                            recordantiguo.status = False
                            recordantiguo.save(update_asignaturamalla=False)

                            if equivalencia.asignaturamallasalto_id in [10822, 10834, 10846, 10858, 10870, 10888, 10913,
                                                                        10930, 10944]:
                                observaciones = recordantiguo.observaciones + " Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15"
                                homologada = True


                            else:
                                observaciones = recordantiguo.observaciones + " Migración con base a RESOLUCIÓN CGA-SO-9-2023-NO15 Y RESOLUCIÓN OCS-SO-18-2023-NO14"
                                homologada = False
                            if not RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                  asignaturamalla=equivalencia.asignaturamallasalto).exists():

                                recordnuevo = RecordAcademico(inscripcion=inscripcion,
                                                              matriculas=recordantiguo.matriculas,
                                                              asignaturamalla=equivalencia.asignaturamallasalto,
                                                              asignatura=equivalencia.asignaturamallasalto.asignatura,
                                                              asignaturaold_id=recordantiguo.asignatura.id,
                                                              nota=recordantiguo.nota,
                                                              asistencia=recordantiguo.asistencia,
                                                              sinasistencia=recordantiguo.sinasistencia,
                                                              fecha=recordantiguo.fecha,
                                                              noaplica=recordantiguo.noaplica,
                                                              aprobada=recordantiguo.aprobada,
                                                              convalidacion=recordantiguo.convalidacion,
                                                              pendiente=recordantiguo.pendiente,
                                                              creditos=equivalencia.asignaturamallasalto.creditos,
                                                              horas=equivalencia.asignaturamallasalto.horas,
                                                              valida=recordantiguo.valida,
                                                              validapromedio=recordantiguo.validapromedio,
                                                              observaciones=observaciones,
                                                              homologada=homologada,
                                                              materiaregular=recordantiguo.materiaregular,
                                                              materiacurso=None,
                                                              completonota=recordantiguo.completonota,
                                                              completoasistencia=recordantiguo.completoasistencia,
                                                              fechainicio=recordantiguo.fechainicio,
                                                              fechafin=recordantiguo.fechafin,
                                                              suficiencia=recordantiguo.suficiencia,
                                                              asignaturamallahistorico_id=recordantiguo.asignaturamalla.id,
                                                              reverso=False)
                                recordnuevo.save()
                                print(u"Crea nuevo record %s" % recordnuevo)


                            elif RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                asignaturamalla=equivalencia.asignaturamallasalto):
                                recordnuevo = RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                             asignaturamalla=equivalencia.asignaturamallasalto)[
                                    0]
                                recordnuevo.matriculas = recordantiguo.matriculas
                                recordnuevo.asignaturamalla = equivalencia.asignaturamallasalto
                                recordnuevo.asignatura = equivalencia.asignaturamallasalto.asignatura
                                recordnuevo.asignaturaold = recordantiguo.asignaturamalla.asignatura
                                recordnuevo.nota = recordantiguo.nota
                                recordnuevo.asistencia = recordantiguo.asistencia
                                recordnuevo.sinasistencia = recordantiguo.sinasistencia
                                recordnuevo.fecha = recordantiguo.fecha
                                recordnuevo.noaplica = recordantiguo.noaplica
                                recordnuevo.aprobada = recordantiguo.aprobada
                                recordnuevo.convalidacion = recordantiguo.convalidacion
                                recordnuevo.pendiente = recordantiguo.pendiente
                                recordnuevo.creditos = equivalencia.asignaturamallasalto.creditos
                                recordnuevo.horas = equivalencia.asignaturamallasalto.horas
                                recordnuevo.valida = recordantiguo.valida
                                recordnuevo.validapromedio = recordantiguo.validapromedio
                                recordnuevo.observaciones = observaciones
                                recordnuevo.homologada = homologada
                                recordnuevo.materiaregular = recordantiguo.materiaregular
                                recordnuevo.materiacurso = None
                                recordnuevo.completonota = recordantiguo.completonota
                                recordnuevo.completoasistencia = recordantiguo.completoasistencia
                                recordnuevo.fechainicio = recordantiguo.fechainicio
                                recordnuevo.fechafin = recordantiguo.fechafin
                                recordnuevo.suficiencia = recordantiguo.suficiencia
                                recordnuevo.asignaturamallahistorico = recordantiguo.asignaturamalla
                                recordnuevo.reverso = False
                                recordnuevo.save()

                            if recordnuevo:
                                historicos = HistoricoRecordAcademico.objects.filter(status=True,
                                                                                     recordacademico=recordantiguo).update(
                                    recordacademico=recordnuevo,
                                    creditos=recordnuevo.creditos,
                                    horas=recordnuevo.horas,
                                    homologada=recordnuevo.homologada)
                                respaldo = RespaldoRecordAcademico.objects.filter(status=True,
                                                                                  recordacademicooriginal=recordantiguo)

                                if equivalencia.asignaturamallasalto_id in [10822, 10834, 10846, 10858, 10870, 10888,
                                                                            10913, 10930, 10944]:
                                    if not practicaspp:
                                        if recordnuevo.aprobada:
                                            profesor = None
                                            if recordnuevo.materiaregular:
                                                profesor = recordnuevo.materiaregular.profesor_principal()
                                            elif recordnuevo.materiacurso:
                                                profesor = recordnuevo.materiaregular.profesor()
                                            if equivalencia.asignaturamallasalto_id == 10822:
                                                itinerarioprimero = ItinerariosMalla.objects.get(status=True,
                                                                                                 malla_id=mallaantigua_id,
                                                                                                 nivel_id=1)
                                                iprimeronuevo = ItinerariosMalla.objects.get(status=True,
                                                                                             malla_id=mallanueva_id,
                                                                                             nivel_id=1)

                                                practica = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud__in=[
                                                        1, 2, 4,
                                                        5, 6],
                                                    itinerariomalla=itinerarioprimero).exists()
                                                practicarechazada = False
                                                if not practica:
                                                    practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        estadosolicitud=3,
                                                        itinerariomalla=itinerarioprimero).exists()

                                                if not practica or practicarechazada:
                                                    if not PracticasPreprofesionalesInscripcion.objects.filter(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            actividad__itinerariomalla=itinerarioprimero).exists():
                                                        nuevapractica = PracticasPreprofesionalesInscripcion(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            numerohora=iprimeronuevo.horas_practicas,
                                                            nivelmalla=iprimeronuevo.nivel,
                                                            tiposolicitud=1,
                                                            estadosolicitud=2,
                                                            tipo=1,
                                                            itinerariomalla=iprimeronuevo,
                                                            supervisor=profesor,
                                                            tutorunemi=profesor,
                                                            fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            tipoinstitucion=1,
                                                            sectoreconomico=6,
                                                            empresaempleadora_id=3,
                                                            culminada=True,
                                                            fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            lugarpractica_id=2,
                                                            observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                        )
                                                        nuevapractica.save()
                                                ItinerariosActividadDetalleDistributivoCarrera.objects.filter(
                                                    status=True,
                                                    itinerario=itinerarioprimero).update(
                                                    itinerario=iprimeronuevo)

                                            if equivalencia.asignaturamallasalto_id == 10834:
                                                itinerariosegundo = ItinerariosMalla.objects.get(status=True,
                                                                                                 malla_id=mallaantigua_id,
                                                                                                 nivel_id=2)
                                                isegundonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                             malla_id=mallanueva_id,
                                                                                             nivel_id=2)

                                                practica = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud__in=[
                                                        1, 2, 4,
                                                        5, 6],
                                                    itinerariomalla=itinerariosegundo).exists()
                                                practicarechazada = False
                                                if not practica:
                                                    practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        estadosolicitud=3,
                                                        itinerariomalla=itinerariosegundo).exists()

                                                if not practica or practicarechazada:
                                                    if not PracticasPreprofesionalesInscripcion.objects.filter(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            actividad__itinerariomalla=itinerariosegundo).exists():
                                                        nuevapractica = PracticasPreprofesionalesInscripcion(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            numerohora=isegundonuevo.horas_practicas,
                                                            nivelmalla=isegundonuevo.nivel,
                                                            tiposolicitud=1,
                                                            estadosolicitud=2,
                                                            tipo=1,
                                                            itinerariomalla=isegundonuevo,
                                                            supervisor=profesor,
                                                            tutorunemi=profesor,
                                                            fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            tipoinstitucion=1,
                                                            sectoreconomico=6,
                                                            empresaempleadora_id=3,
                                                            culminada=True,
                                                            fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            lugarpractica_id=2,
                                                            observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                        )
                                                        nuevapractica.save()
                                                ItinerariosActividadDetalleDistributivoCarrera.objects.filter(
                                                    status=True,
                                                    itinerario=itinerariosegundo).update(
                                                    itinerario=isegundonuevo)

                                            if equivalencia.asignaturamallasalto_id == 10846:
                                                itinerariotercero = ItinerariosMalla.objects.get(status=True,
                                                                                                 malla_id=mallaantigua_id,
                                                                                                 nivel_id=3)
                                                iterceronuevo = ItinerariosMalla.objects.get(status=True,
                                                                                             malla_id=mallanueva_id,
                                                                                             nivel_id=3)

                                                practica = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud__in=[
                                                        1, 2, 4,
                                                        5, 6],
                                                    itinerariomalla=itinerariotercero).exists()
                                                practicarechazada = False
                                                if not practica:
                                                    practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        estadosolicitud=3,
                                                        itinerariomalla=itinerariotercero).exists()

                                                if not practica or practicarechazada:
                                                    if not PracticasPreprofesionalesInscripcion.objects.filter(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            actividad__itinerariomalla=itinerariotercero).exists():
                                                        nuevapractica = PracticasPreprofesionalesInscripcion(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            numerohora=iterceronuevo.horas_practicas,
                                                            nivelmalla=iterceronuevo.nivel,
                                                            tiposolicitud=1,
                                                            estadosolicitud=2,
                                                            tipo=1,
                                                            itinerariomalla=iterceronuevo,
                                                            supervisor=profesor,
                                                            tutorunemi=profesor,
                                                            fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            tipoinstitucion=1,
                                                            sectoreconomico=6,
                                                            empresaempleadora_id=3,
                                                            culminada=True,
                                                            fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            lugarpractica_id=2,
                                                            observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                        )
                                                        nuevapractica.save()

                                                ItinerariosActividadDetalleDistributivoCarrera.objects.filter(
                                                    status=True,
                                                    itinerario=itinerariotercero).update(
                                                    itinerario=iterceronuevo)

                                            if equivalencia.asignaturamallasalto_id == 10858:
                                                itinerariocuarto = ItinerariosMalla.objects.get(status=True,
                                                                                                malla_id=mallaantigua_id,
                                                                                                nivel_id=4)
                                                icuartonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                            malla_id=mallanueva_id,
                                                                                            nivel_id=4)

                                                practica = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud__in=[
                                                        1, 2, 4,
                                                        5, 6],
                                                    itinerariomalla=itinerariocuarto).exists()
                                                practicarechazada = False
                                                if not practica:
                                                    practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        estadosolicitud=3,
                                                        itinerariomalla=itinerariocuarto).exists()

                                                if not practica or practicarechazada:
                                                    if not PracticasPreprofesionalesInscripcion.objects.filter(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            actividad__itinerariomalla=itinerariocuarto).exists():
                                                        nuevapractica = PracticasPreprofesionalesInscripcion(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            numerohora=icuartonuevo.horas_practicas,
                                                            nivelmalla=icuartonuevo.nivel,
                                                            tiposolicitud=1,
                                                            estadosolicitud=2,
                                                            tipo=1,
                                                            itinerariomalla=icuartonuevo,
                                                            supervisor=profesor,
                                                            tutorunemi=profesor,
                                                            fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            tipoinstitucion=1,
                                                            sectoreconomico=6,
                                                            empresaempleadora_id=3,
                                                            culminada=True,
                                                            fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            lugarpractica_id=2,
                                                            observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                        )
                                                        nuevapractica.save()
                                                ItinerariosActividadDetalleDistributivoCarrera.objects.filter(
                                                    status=True,
                                                    itinerario=itinerariocuarto).update(
                                                    itinerario=icuartonuevo)

                                            if equivalencia.asignaturamallasalto_id == 10870:
                                                itinerarioquinto = ItinerariosMalla.objects.get(status=True,
                                                                                                malla_id=mallaantigua_id,
                                                                                                nivel_id=5)
                                                iquintonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                            malla_id=mallanueva_id,
                                                                                            nivel_id=5)

                                                practica = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud__in=[
                                                        1, 2, 4,
                                                        5, 6],
                                                    itinerariomalla=itinerarioquinto).exists()
                                                practicarechazada = False
                                                if not practica:
                                                    practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        estadosolicitud=3,
                                                        itinerariomalla=itinerarioquinto).exists()

                                                if not practica or practicarechazada:
                                                    if not PracticasPreprofesionalesInscripcion.objects.filter(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            actividad__itinerariomalla=itinerarioquinto).exists():
                                                        nuevapractica = PracticasPreprofesionalesInscripcion(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            numerohora=iquintonuevo.horas_practicas,
                                                            nivelmalla=iquintonuevo.nivel,
                                                            tiposolicitud=1,
                                                            estadosolicitud=2,
                                                            tipo=1,
                                                            itinerariomalla=iquintonuevo,
                                                            supervisor=profesor,
                                                            tutorunemi=profesor,
                                                            fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            tipoinstitucion=1,
                                                            sectoreconomico=6,
                                                            empresaempleadora_id=3,
                                                            culminada=True,
                                                            fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            lugarpractica_id=2,
                                                            observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                        )
                                                        nuevapractica.save()

                                                ItinerariosActividadDetalleDistributivoCarrera.objects.filter(
                                                    status=True,
                                                    itinerario=itinerarioquinto).update(
                                                    itinerario=iquintonuevo)

                                            if equivalencia.asignaturamallasalto_id == 10888:
                                                itinerariosexto = ItinerariosMalla.objects.get(status=True,
                                                                                               malla_id=mallaantigua_id,
                                                                                               nivel_id=6)
                                                isextonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallanueva_id,
                                                                                           nivel_id=6)

                                                practica = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud__in=[
                                                        1, 2, 4,
                                                        5, 6],
                                                    itinerariomalla=itinerariosexto).exists()
                                                practicarechazada = False
                                                if not practica:
                                                    practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        estadosolicitud=3,
                                                        itinerariomalla=itinerariosexto).exists()

                                                if not practica or practicarechazada:
                                                    if not PracticasPreprofesionalesInscripcion.objects.filter(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            actividad__itinerariomalla=itinerariosexto).exists():
                                                        nuevapractica = PracticasPreprofesionalesInscripcion(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            numerohora=isextonuevo.horas_practicas,
                                                            nivelmalla=isextonuevo.nivel,
                                                            tiposolicitud=1,
                                                            estadosolicitud=2,
                                                            tipo=1,
                                                            itinerariomalla=isextonuevo,
                                                            supervisor=profesor,
                                                            tutorunemi=profesor,
                                                            fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            tipoinstitucion=1,
                                                            sectoreconomico=6,
                                                            empresaempleadora_id=3,
                                                            culminada=True,
                                                            fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            lugarpractica_id=2,
                                                            observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                        )
                                                        nuevapractica.save()
                                                ItinerariosActividadDetalleDistributivoCarrera.objects.filter(
                                                    status=True,
                                                    itinerario=itinerariosexto).update(
                                                    itinerario=isextonuevo)

                                            if equivalencia.asignaturamallasalto_id == 10913:
                                                itinerarioseptimo = ItinerariosMalla.objects.get(status=True,
                                                                                                 malla_id=mallaantigua_id,
                                                                                                 nivel_id=7)
                                                iseptimonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                             malla_id=mallanueva_id,
                                                                                             nivel_id=7)
                                                practica = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud__in=[
                                                        1, 2, 4,
                                                        5, 6],
                                                    itinerariomalla=itinerarioseptimo).exists()
                                                practicarechazada = False
                                                if not practica:
                                                    practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        estadosolicitud=3,
                                                        itinerariomalla=itinerarioseptimo).exists()

                                                if not practica or practicarechazada:
                                                    if not PracticasPreprofesionalesInscripcion.objects.filter(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            actividad__itinerariomalla=itinerarioseptimo).exists():
                                                        nuevapractica = PracticasPreprofesionalesInscripcion(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            numerohora=iseptimonuevo.horas_practicas,
                                                            nivelmalla=iseptimonuevo.nivel,
                                                            tiposolicitud=1,
                                                            estadosolicitud=2,
                                                            tipo=1,
                                                            itinerariomalla=iseptimonuevo,
                                                            supervisor=profesor,
                                                            tutorunemi=profesor,
                                                            fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            tipoinstitucion=1,
                                                            sectoreconomico=6,
                                                            empresaempleadora_id=3,
                                                            culminada=True,
                                                            fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            lugarpractica_id=2,
                                                            observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                        )
                                                        nuevapractica.save()
                                                ItinerariosActividadDetalleDistributivoCarrera.objects.filter(
                                                    status=True,
                                                    itinerario=itinerarioseptimo).update(
                                                    itinerario=iseptimonuevo)

                                            if equivalencia.asignaturamallasalto_id == 10930:
                                                itinerariooctavo = ItinerariosMalla.objects.get(status=True,
                                                                                                malla_id=mallaantigua_id,
                                                                                                nivel_id=8)
                                                ioctavonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                            malla_id=mallanueva_id,
                                                                                            nivel_id=8)

                                                practica = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud__in=[
                                                        1, 2, 4,
                                                        5, 6],
                                                    itinerariomalla=itinerariooctavo).exists()
                                                practicarechazada = False
                                                if not practica:
                                                    practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        estadosolicitud=3,
                                                        itinerariomalla=itinerariooctavo).exists()

                                                if not practica or practicarechazada:
                                                    if not PracticasPreprofesionalesInscripcion.objects.filter(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            actividad__itinerariomalla=itinerariooctavo).exists():
                                                        nuevapractica = PracticasPreprofesionalesInscripcion(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            numerohora=ioctavonuevo.horas_practicas,
                                                            nivelmalla=ioctavonuevo.nivel,
                                                            tiposolicitud=1,
                                                            estadosolicitud=2,
                                                            tipo=1,
                                                            itinerariomalla=ioctavonuevo,
                                                            supervisor=profesor,
                                                            tutorunemi=profesor,
                                                            fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            tipoinstitucion=1,
                                                            sectoreconomico=6,
                                                            empresaempleadora_id=3,
                                                            culminada=True,
                                                            fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            lugarpractica_id=2,
                                                            observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                        )
                                                        nuevapractica.save()
                                                ItinerariosActividadDetalleDistributivoCarrera.objects.filter(
                                                    status=True,
                                                    itinerario=itinerariooctavo).update(
                                                    itinerario=ioctavonuevo)

                                            if equivalencia.asignaturamallasalto_id == 10944:
                                                itinerarionoveno = ItinerariosMalla.objects.get(status=True,
                                                                                                malla_id=mallaantigua_id,
                                                                                                nivel_id=9)
                                                inovenonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                            malla_id=mallanueva_id,
                                                                                            nivel_id=9)
                                                practica = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud__in=[
                                                        1, 2, 4,
                                                        5, 6],
                                                    itinerariomalla=itinerarionoveno).exists()
                                                practicarechazada = False
                                                if not practica:
                                                    practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        estadosolicitud=3,
                                                        itinerariomalla=itinerarionoveno).exists()

                                                if not practica or practicarechazada:
                                                    if not PracticasPreprofesionalesInscripcion.objects.filter(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            actividad__itinerariomalla=itinerarionoveno).exists():
                                                        nuevapractica = PracticasPreprofesionalesInscripcion(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            numerohora=inovenonuevo.horas_practicas,
                                                            nivelmalla=inovenonuevo.nivel,
                                                            tiposolicitud=1,
                                                            estadosolicitud=2,
                                                            tipo=1,
                                                            itinerariomalla=inovenonuevo,
                                                            supervisor=profesor,
                                                            tutorunemi=profesor,
                                                            fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            tipoinstitucion=1,
                                                            sectoreconomico=6,
                                                            empresaempleadora_id=3,
                                                            culminada=True,
                                                            fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            lugarpractica_id=2,
                                                            observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                        )
                                                        nuevapractica.save()
                                                ItinerariosActividadDetalleDistributivoCarrera.objects.filter(
                                                    status=True,
                                                    itinerario=itinerarionoveno).update(
                                                    itinerario=inovenonuevo)

                                if equivalencia.asignaturamallasalto_id in [10913, 10930]:
                                    if not horasvinculacion:
                                        if recordnuevo.aprobada:
                                            if equivalencia.asignaturamallasalto_id == 10913 and inscripcion.numero_horas_proyectos_vinculacion() < 80:
                                                horasfalta = 80 - inscripcion.numero_horas_proyectos_vinculacion()
                                                vinculacion = ParticipantesMatrices(status=True,
                                                                                    matrizevidencia_id=2,
                                                                                    proyecto_id=601,
                                                                                    inscripcion=inscripcion,
                                                                                    horas=horasfalta,
                                                                                    registrohorasdesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                    registrohorashasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                    estado=1
                                                                                    )
                                                vinculacion.save()

                                            if equivalencia.asignaturamallasalto_id == 10930 and inscripcion.numero_horas_proyectos_vinculacion() < 160:
                                                horasfalta = 160 - inscripcion.numero_horas_proyectos_vinculacion()
                                                vinculacion = ParticipantesMatrices(status=True,
                                                                                    matrizevidencia_id=2,
                                                                                    proyecto_id=601,
                                                                                    inscripcion=inscripcion,
                                                                                    horas=horasfalta,
                                                                                    registrohorasdesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                    registrohorashasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                                                    estado=1
                                                                                    )
                                                vinculacion.save()

                                if not respaldo.exists():
                                    respaldorecord = RespaldoRecordAcademico(
                                        recordacademicooriginal=recordantiguo,
                                        recordacademiconuevo=recordnuevo
                                    )
                                    respaldorecord.save()
                                else:
                                    respaldorecord = respaldo[0]
                                    respaldorecord.recordacademiconuevo = recordnuevo
                                    respaldorecord.save()
                                print(u"Record actualizado %s" % recordnuevo)

                        else:
                            hojadestino.write(fila, 3, "NO ENCONTRO RECORD ANTIGUO %s" % equivalencia.asignaturamalla)
                            fila += 1

                    # if cont_asig_vinculacion_aprobadas != 0:
                    #     if cont_asig_vinculacion_aprobadas == 1:
                    #         horasfalta = 80
                    #     elif cont_asig_vinculacion_aprobadas == 2:
                    #         horasfalta = 160
                    #     horasfalta = horasfalta - inscripcion.numero_horas_proyectos_vinculacion()
                    #     vinculacion = ParticipantesMatrices(status=True,
                    #                                         matrizevidencia_id=2,
                    #                                         proyecto_id=601,
                    #                                         inscripcion=inscripcion,
                    #                                         horas=horasfalta,
                    #                                         registrohorasdesde=fechainicioitinerario,
                    #                                         registrohorashasta=fechafinitinerario,
                    #                                         estado=1
                    #                                         )
                    #     vinculacion.save()

                    practicasppf = inscripcion.numero_horas_practicas_pre_profesionales()
                    hojadestino.write(fila, 3, practicasppf, fuentenormal)
                    horasvinculacionf = inscripcion.numero_horas_proyectos_vinculacion()
                    hojadestino.write(fila, 4, horasvinculacionf, fuentenormal)
                    fila += 1

                    time.sleep(1)

            lin += 1

        libdestino.save(output_folder + libre_origen)
        print(output_folder + libre_origen)
        print("Proceso finalizado. . .")

    except Exception as ex:
        transaction.set_rollback(True)
        print('error: %s' % ex)
        hojadestino.write(fila, 3, str(ex))
        fila += 1

homologacion_basica2()