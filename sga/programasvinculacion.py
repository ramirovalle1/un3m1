# -*- coding: UTF-8 -*-
import io
import os
import xlsxwriter
import random
import json
from datetime import datetime, date, timedelta

import sys
import xlwt
from django.core.exceptions import ObjectDoesNotExist
# from openpyxl.reader.excel import load_workbook
import openpyxl
from openpyxl import load_workbook
from copy import copy
from openpyxl.styles import Border, Side
from openpyxl.styles.alignment import Alignment as alin
from settings import  MEDIA_ROOT

from investigacion.forms import DocenteExternoForm
from sga.tasks import send_html_mail
from django.db.models import Q, Sum, Count
from django.contrib.auth.decorators import login_required
from django.db import transaction, connection
from django.db.models.aggregates import Sum
from django.http import JsonResponse, request
from django.http.response import HttpResponse, HttpResponseRedirect
from django.shortcuts import render
from django.template.context import Context
from django.template.loader import get_template
from xlwt import *
from decorators import secure_module, last_access
from sga.commonviews import adduserdata, traerNotificaciones
from sga.templatetags.sga_extras import encrypt
import io
import xlsxwriter
from django.forms import  model_to_dict
from sga.forms import ProgramasInvestigacionForm, EvidenciaForm, ProyectoInvestigacionForm, \
    ParticipanteProfesorForm, ParticipanteEstudianteForm, ParticipanteAdministrativoForm, PresupuestoProyectoForm, \
    ProgramasVinculacionForm, AprobacionProyecto, ProyectoVinculacion1Form, EvidenciaFormatoForm, \
    ProgramasVinculacionLineaForm, \
    ProgramasVinculacionAreaForm, FechaProyectosFrom, FechaInformeForm, FechaFinProyectosFrom, FechaFinProyectoFrom, \
    CargoBeneficiarioFrom, ResolucionForm, CarreraParticipanteForm, PerfilProfesionalForm, BeneficiariosForm, \
    InvolucradoForm, ArProblemaForm, ArProb_CausaEfectoForm, DatoSecundarioForm, LineaBaseForm, ArObjetivoForm, \
    MarcoLogicoForm2, \
    CronogramaForm, PresupuestoForm, RedaccionForm, AnexosForm, ParticipanteProfesorForm2, FechaEntregaProyectoFrom, \
    AprobacionCumplimientoForm, AprobacionInformeForm, ConfiguracionCambioForm, FechaGestionForm,CambioDocenteVinculacionForm,\
    SubirInformeForm,PeriodoInscripcionFrom,CambioConvocatoriaForm, FechasRegistroHorasForm, InformeProyectoVinculacionForm,\
    CambioInscripcionVinculacionForm, HabilitacionRegistroHorasVinculacionForm,SuministroForm, InscripcionManualProyectoVinculacionForm,\
    CambioCarreraVinculacionForm,RegistrarHorasVinculacionForm, ParticipanteProfesorVinculacionForm, MasivoInscripcionVinculacionForm
from sga.funciones import log, generar_nombre, MiPaginador,null_to_decimal, get_director_vinculacion, variable_valor,cuenta_email_disponible
from inno.funciones import haber_cumplido_horas_creditos_vinculacion, obtener_materia_asignada_vinculacion_por_nivel_v2
from sga.funcionesxhtml2pdf import conviert_html_to_pdf
from sga.models import Inscripcion, ProgramasInvestigacion, DetalleEvidencias, ProyectosInvestigacion, \
    ParticipantesMatrices, PresupuestosProyecto, CarrerasProyecto, Graduado, Evidencia, ParticipantesTipo, Carrera, \
    ProgramasVinculacionCampos, PropuestaSubLineaInvestigacion, ProyectoVinculacionCampos, Provincia, Canton, \
    SubLineaInvestigacion, LineaInvestigacion, LineaProgramasInvestigacion, PeriodoMatriculacion, \
    AreaProgramasInvestigacion, LineamientoPlanNacional, MetasComplementariaPlanNacional, \
    MetasComplementariaProyectosInvestigacion, LineamientoProyectosInvestigacion, PoliticaProyectosInvestigacion, \
    PoliticasObjetivoPlanNacional, MetasPndProyectosInvestigacion, MetasZonalPlanNacional, \
    MetasZonalProyectosInvestigacion, MetasPndPlanNacional, AreaConocimientoTitulacion, ProyectosInvestigacionCarreras, \
    FechaProyectos, SubAreaConocimientoTitulacion, SubAreaEspecificaConocimientoTitulacion, DetalleInforme, \
    InformeMarcoLogicoProyectosInvestigacion, ProyectosInvestigacionAprobacion, ProyectosInvestigacionZonas, \
    ProyectosInvestigacionCantones, Zona, ProyectoVinculacionInscripcion, CargoBeneficiario, Resolucion, \
    MatrizLineaBase, Beneficiarios, Involucrado, Presupuesto, CarrerasParticipantes, DatoSecundario, ArbolObjetivo, \
    ArbolProblema, Cronograma, Anexos, MarcoLogico, PerfilProfesional, Problema, Profesor, \
    ESTADOS_PROYECTO_VINCULACION_INVESTIGACION, Periodo, DetalleCumplimiento, ConfiguracionInformeVinculacion, \
    DetalleInformeVinculacion, MarcoLogicoReporte, ConfiguracionCambio, PeriodoInscripcionVinculacion, CarreraInscripcionVinculacion, \
    InformesProyectoVinculacionDocente, miinstitucion, InscripcionActividadConvalidacionPPV, HabilitacionesHorasParticipantesVinculacion, \
    Persona, Suministro, InformesProyectoVinculacionEstudiante, ItinerariosVinculacionMalla, CUENTAS_CORREOS, Matricula, Notificacion, Externo, \
    CriterioDocencia, DetalleDistributivo
from sagest.models import Producto, DetalleIngresoProducto
from django.db.models.functions import ExtractYear
from django.contrib import messages
from core.firmar_documentos import firmar, obtener_posicion_x_y_saltolinea, verificarFirmasPDF
from sga.excelbackground import reporte_estudiantes_matriculados_vinculacion_background, reporte_estudiantes_proceso_vinculacion_background
from core.firmar_documentos_ec import JavaFirmaEc
from postulaciondip.forms import FirmaElectronicaIndividualForm
from django.core.files import File as DjangoFile
from django.core.files.base import ContentFile
from inno.models import TecnicoAsociadoProyectoVinculacion, ExtraProyectoVinculacionInscripcion
from inno.forms import TecnicoAsociadoProyectoVinculacionForm
from sga.proyectovinculaciondocente import migrar_evidencia_proyecto_vinculacion
from settings import DEBUG, SITE_STORAGE

@login_required(redirect_field_name='ret', login_url='/loginsga')
@secure_module
@last_access
@transaction.atomic()
def view(request):
    data = {}
    adduserdata(request, data)
    persona = request.session['persona']
    perfilprincipal = request.session['perfilprincipal']
    responsablevinculacion = get_director_vinculacion()
    miscarreras = persona.mis_carreras()
    cargosactuales = persona.mis_cargos_actuales().values_list('id', flat=True)
    experto_presu = data["grupos_usuarios"].filter(id__in=variable_valor('VALIDACION_GRUPO_PROYECTO_VINCULACION')).exists()

    # data['periodo'] = periodo = request.session['periodo']
    periodo = request.session['periodo']

    dominio_sistema = 'https://sga.unemi.edu.ec'
    if DEBUG: dominio_sistema = 'http://localhost:8000'

    data["DOMINIO_DEL_SISTEMA"] = dominio_sistema
    if request.method == 'POST':
        action = request.POST['action']
        if action == 'add':
            try:
                f = ProgramasVinculacionForm(request.POST)
                if f.is_valid():
                    if not ProgramasInvestigacion.objects.filter(nombre=f.cleaned_data['nombre']).exists():
                        programas = ProgramasInvestigacion(nombre=f.cleaned_data['nombre'],
                                                           fechainicio=f.cleaned_data['fechainicio'],
                                                           fechaplaneado=f.cleaned_data['fechaplaneado'],
                                                           fechareal=f.cleaned_data['fechareal'],
                                                           # lineainvestigacion=f.cleaned_data['lineainvestigacion'],
                                                           alcanceterritorial=f.cleaned_data['alcanceterritorial'],
                                                           # areaconocimiento=f.cleaned_data['areaconocimiento'],
                                                           # subareaconocimiento=f.cleaned_data['subareaconocimiento'],
                                                           # subareaespecificaconocimiento=f.cleaned_data['subareaespecificaconocimiento'],
                                                           tiempoejecucion=f.cleaned_data['tiempoejecucion'],
                                                           # sublineainvestigacion=f.cleaned_data['sublineainvestigacion'],
                                                           valorpresupuestointerno=f.cleaned_data['valorpresupuestointerno'],
                                                           valorpresupuestoexterno=f.cleaned_data['valorpresupuestoexterno']
                                                           )
                        programas.save(request)
                        programasvinculacioncampos = ProgramasVinculacionCampos(programasvinculacion=programas,
                                                                                soportedelprograma=f.cleaned_data['soportedelprograma'],
                                                                                elaboradopor=f.cleaned_data['elaboradopor'],
                                                                                perfilbeneficiarios=f.cleaned_data['perfilbeneficiarios'],
                                                                                numerobeneficiarios=f.cleaned_data['numerobeneficiarios'],
                                                                                proyectosintegranprograma=f.cleaned_data['proyectosintegranprograma'],
                                                                                perfildocentesestudiantes=f.cleaned_data['perfildocentesestudiantes'],
                                                                                planteamientoproblema=f.cleaned_data['planteamientoproblema'],
                                                                                justificacion=f.cleaned_data['justificacion'],
                                                                                pertinencia=f.cleaned_data['pertinencia'],
                                                                                objetivogeneral=f.cleaned_data['objetivogeneral'],
                                                                                objetivoespecifico=f.cleaned_data['objetivoespecifico'],
                                                                                metodologia=f.cleaned_data['metodologia'],
                                                                                recursos=f.cleaned_data['recursos'],
                                                                                cronograma=f.cleaned_data['cronograma'],
                                                                                seguimiento=f.cleaned_data['seguimiento'],
                                                                                evaluacion=f.cleaned_data['evaluacion'],
                                                                                bibliografia=f.cleaned_data['bibliografia'],
                                                                                anexos=f.cleaned_data['anexos'])
                        programasvinculacioncampos.save(request)
                        log(u'Adiciono programa de investigación: %s' % programas, request, "add")
                        return JsonResponse({"result": "ok"})
                    else:
                        return JsonResponse({"result": "bad", "mensaje": u"Ya el estudiante esta graduado."})
                else:
                     raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        if action == 'addproyecto':
            try:
                f = ProyectoVinculacion1Form(request.POST)
                if f.is_valid():
                    if not ProyectosInvestigacion.objects.filter(nombre=f.cleaned_data['nombre']).exists():
                        if not request.POST['conv'] =="":
                            conv = FechaProyectos.objects.get(pk=request.POST['conv'])
                            estado = 5
                        else:
                            conv=None
                            estado = 6
                        proyecto = ProyectosInvestigacion(
                                                        convocatoria=conv,
                                                        nombre=f.cleaned_data['nombre'],
                                                        fechainicio=f.cleaned_data['fechainicio'],
                                                        fechaplaneacion=f.cleaned_data['fechaPlanificacion'],
                                                        programa=f.cleaned_data['programa'],
                                                        tipo=1,
                                                        tipoproinstitucion=f.cleaned_data['tipoproinstitucion'],
                                                        alcanceterritorial=f.cleaned_data['alcanceterritorial'],
                                                        areaconocimiento=f.cleaned_data['areaconocimiento'],
                                                        subareaconocimiento=f.cleaned_data['subareaconocimiento'],
                                                        subareaespecificaconocimiento=f.cleaned_data['subareaespecificaconocimiento'],
                                                        tiempoejecucion=f.cleaned_data['tiempoejecucion'],
                                                        tiempo_duracion_horas=f.cleaned_data['tiempo_duracion_horas'],
                                                        lineainvestigacion=f.cleaned_data['lineainvestigacion'],
                                                        sublineainvestigacion=f.cleaned_data['sublineainvestigacion'],
                                                        sectorcoordenada=f.cleaned_data['sectorcoordenada'],
                                                        objetivos_PND=f.cleaned_data['objetivos_PND'],
                                                        politicas_PND=f.cleaned_data['politicas_PND'],
                                                        linea_accion=f.cleaned_data['linea_accion'],
                                                        estrategia_desarrollo=f.cleaned_data['estrategia_desarrollo'],
                                                        investigacion_institucional=f.cleaned_data['investigacion_institucional'],
                                                        necesidades_sociales=f.cleaned_data['necesidades_sociales'],
                                                        circuito=f.cleaned_data['circuito'],
                                                        distrito=f.cleaned_data['distrito'],
                                                        # canton=f.cleaned_data['canton'],
                                                        # zona=f.cleaned_data['zona'],
                                                        aprobacion=estado,
                        )

                        proyecto.save(request)
                        # proyecto.canton = f.cleaned_data["canton"]
                        # proyecto.zona = f.cleaned_data["zona"]
                        for cant in f.cleaned_data['canton']:
                            proyecto.canton.add(cant)
                        for z in f.cleaned_data['zona']:
                            proyecto.zona.add(z)

                        log(u'Adiciono proyecto de investigación docente: %s' % proyecto, request, "add")
                        return JsonResponse({"result": "ok"})
                    else:
                        return JsonResponse({"result": "bad", "mensaje": u"Error a guardar."})
                else:
                     raise()
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'editproyecto':
            try:
                f = ProyectoVinculacion1Form(request.POST)
                proyecto = ProyectosInvestigacion.objects.select_related().get(pk=request.POST['id'],tipo=1)
                if f.is_valid():
                    # bandera = False
                    # newfile = None
                    #
                    # if proyecto.objetivoplannacional:
                    # if proyecto.objetivoplannacional:
                    #     if proyecto.objetivoplannacional.id == f.cleaned_data['objetivoplannacional'].id:
                    #         bandera = True


                    proyecto.nombre = f.cleaned_data['nombre']
                    proyecto.fechainicio = f.cleaned_data['fechainicio']
                    # proyecto.fechafin=f.cleaned_data['fechafin']
                    proyecto.fechaplaneacion = f.cleaned_data['fechaPlanificacion']
                    proyecto.programa = f.cleaned_data['programa']
                    proyecto.tipo = 1
                    proyecto.tipoproinstitucion = f.cleaned_data['tipoproinstitucion']
                    proyecto.alcanceterritorial = f.cleaned_data['alcanceterritorial']
                    proyecto.areaconocimiento = f.cleaned_data['areaconocimiento']
                    proyecto.subareaconocimiento = f.cleaned_data['subareaconocimiento']
                    proyecto.subareaespecificaconocimiento = f.cleaned_data['subareaespecificaconocimiento']
                    proyecto.tiempoejecucion = f.cleaned_data['tiempoejecucion']
                    proyecto.lineainvestigacion = f.cleaned_data['lineainvestigacion']
                    proyecto.sublineainvestigacion = f.cleaned_data['sublineainvestigacion']
                    # proyecto.valorpresupuestointerno = f.cleaned_data['valorpresupuestointerno']
                    # proyecto.valorpresupuestoexterno = f.cleaned_data['valorpresupuestoexterno']
                    proyecto.sectorcoordenada = f.cleaned_data['sectorcoordenada']
                    # proyecto.presupuestototal = f.cleaned_data['valorpresupuestointerno'] + f.cleaned_data['valorpresupuestoexterno']
                    # proyecto.institucionbeneficiaria = f.cleaned_data['institucionbeneficiaria']
                    # proyecto.objetivoplannacional = f.cleaned_data['objetivoplannacional']
                    # proyecto.cupo = f.cleaned_data['cupo']
                    # proyecto.aprobacion = 5
                    # proyecto.saldoo = f.cleaned_data['cupo']

                    proyecto.objetivos_PND =  f.cleaned_data['objetivos_PND']
                    proyecto.politicas_PND = f.cleaned_data['politicas_PND']
                    proyecto.linea_accion = f.cleaned_data['linea_accion']
                    proyecto.estrategia_desarrollo = f.cleaned_data['estrategia_desarrollo']
                    proyecto.investigacion_institucional = f.cleaned_data['investigacion_institucional']
                    proyecto.necesidades_sociales = f.cleaned_data['necesidades_sociales']
                    proyecto.tiempo_duracion_horas = f.cleaned_data['tiempo_duracion_horas']
                    proyecto.circuito = f.cleaned_data['circuito']
                    proyecto.distrito = f.cleaned_data['distrito']
                    proyecto.save(request)
                    # proyecto.canton = f.cleaned_data['canton']
                    # proyecto.zona = f.cleaned_data['zona']
                    proyecto.canton.clear()
                    for cant in f.cleaned_data['canton']:
                        proyecto.canton.add(cant)
                    proyecto.zona.clear()
                    for z in f.cleaned_data['zona']:
                        proyecto.zona.add(z)
                    request.session['pestaña'] = 'datos'
                    request.session['vista'] = '1'
                    # proyecto.carrera = f.cleaned_data['carreras']
                    # lineas= f.cleaned_data['carreras']
                    # for lis in lineas:
                    #     detalle=ProyectosInvestigacionCarreras.objects.get(proyectovinculacion=proyecto,carreras_id=lis.id)
                    #     detalle.carreras=Carrera.objects.get(pk=int(lis.id))
                    #     detalle.save(request)



                    # cantones = f.cleaned_data['canton']

                    # for zona in zonas:
                    #     proyectozona = ProyectosInvestigacionZonas.objects.get(proyectovinculacion__id=proyecto.id, zona_id=zona.id)
                    #     proyectozona.proyectovinculacion=proyecto
                    #     proyectozona.zona=zona
                    #     proyectozona.save(request)
                    # for canton in cantones:
                    #     proyectocanton = ProyectosInvestigacionCantones.objects.get(proyectovinculacion__id=proyecto.id,
                    #                                                     canton_id=canton.id)
                    #     proyectocanton.canton=canton
                    #     proyectocanton.proyectovinculacion=proyecto
                    #     proyectocanton.save(request)

                    # if not bandera:
                    #     PoliticaProyectosInvestigacion.objects.filter(proyectovinculacion=proyecto).delete()
                    #     MetasPndProyectosInvestigacion.objects.filter(proyectovinculacion=proyecto).delete()
                    #     MetasZonalProyectosInvestigacion.objects.filter(proyectovinculacion=proyecto).delete()
                    #     MetasComplementariaProyectosInvestigacion.objects.filter(proyectovinculacion=proyecto).delete()
                    #     LineamientoProyectosInvestigacion.objects.filter(proyectovinculacion=proyecto).delete()
                    #     for p in PoliticasObjetivoPlanNacional.objects.filter(objetivo=f.cleaned_data['objetivoplannacional'], status=True):
                    #         politicaproyectosinvestigacion = PoliticaProyectosInvestigacion(proyectovinculacion=proyecto,
                    #                                                                         politica=p,
                    #                                                                         status=False)
                    #         politicaproyectosinvestigacion.save(request)
                    #     # metaspnd
                    #     for p in MetasPndPlanNacional.objects.filter(objetivo=f.cleaned_data['objetivoplannacional'], status=True):
                    #         metaspndproyectosinvestigacion = MetasPndProyectosInvestigacion(proyectovinculacion=proyecto,
                    #                                                                         metaspnd=p,
                    #                                                                         status=False)
                    #         metaspndproyectosinvestigacion.save(request)
                    #     # metaszona
                    #     for p in MetasZonalPlanNacional.objects.filter(objetivo=f.cleaned_data['objetivoplannacional'], status=True):
                    #         metaszonalproyectosinvestigacion = MetasZonalProyectosInvestigacion(proyectovinculacion=proyecto,
                    #                                                                             metaszonal=p,
                    #                                                                             status=False)
                    #         metaszonalproyectosinvestigacion.save(request)
                    #     # metascomplementaria
                    #     for p in MetasComplementariaPlanNacional.objects.filter(objetivo=f.cleaned_data['objetivoplannacional'], status=True):
                    #         metascomplementariaproyectosinvestigacion = MetasComplementariaProyectosInvestigacion(proyectovinculacion=proyecto,
                    #                                                                                               metascomplementaria=p,
                    #                                                                                               status=False)
                    #         metascomplementariaproyectosinvestigacion.save(request)
                    #     # lineamiento
                    #     for p in LineamientoPlanNacional.objects.filter(objetivo=f.cleaned_data['objetivoplannacional'], status=True):
                    #         lineamientoproyectosinvestigacion = LineamientoProyectosInvestigacion(proyectovinculacion=proyecto,
                    #                                                                               lineamiento=p,
                    #                                                                               status=False)
                    #         lineamientoproyectosinvestigacion.save(request)

                    log(u'Editó proyecto de investigación docente: %s' % proyecto, request, "del")
                    return JsonResponse({"result": "ok"})
                else:
                     raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'addcarreraparticipante':
            try:
                f = CarreraParticipanteForm(request.POST)
                if f.is_valid():
                    if not CarrerasParticipantes.objects.filter(proyecto__pk = request.POST['idproyecto'], carrera = f.cleaned_data['carrera'], status=True ).exists():
                        carrera = CarrerasParticipantes(
                                                        proyecto=ProyectosInvestigacion.objects.get(pk=request.POST['idproyecto']),
                                                        carrera= f.cleaned_data['carrera'],
                                                        # cupos= f.cleaned_data['cupos']
                        )
                        carrera.save(request)
                        request.session['pestaña'] = 'carreraParticipante'
                        request.session['vista'] = '3'
                        log(u'Adiciono carrera al proyecto: %s' % carrera, request, "add")
                        return JsonResponse({"result": "ok"})
                    else:
                        return JsonResponse({"result": "bad", "mensaje": u"Carrera ya se encuentra registrada en el proyecto"})
                else:
                     raise()
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'editcarreraparticipante':
            try:
                f = CarreraParticipanteForm(request.POST)
                if f.is_valid():
                    datos = CarrerasParticipantes.objects.get(pk = request.POST['id'])
                    datos.carrera = f.cleaned_data['carrera']
                    # datos.cupos = f.cleaned_data['cupos']
                    datos.save(request)
                    request.session['pestaña'] = 'carreraParticipante'
                    request.session['vista'] = '3'
                    log(u'Edito carrera al proyecto: %s' % datos, request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                     raise()
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'deletecarreraparticipante':
            try:
                participante = CarrerasParticipantes.objects.get(pk=request.POST['id'])
                participante.status = False

                if PerfilProfesional.objects.filter(carreras_participantes = participante, status = True).exists():
                    perfil = PerfilProfesional.objects.get(carreras_participantes = participante, status = True)
                    perfil.status = False
                    perfil.save(request)
                log(u'Elimino carrera practicipante de investigación docente: %s' % participante, request, "del")
                participante.save(request)
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        elif action == 'addperfilprofesional':
            try:
                f = PerfilProfesionalForm(request.POST)
                if f.is_valid():
                    perfil = PerfilProfesional(
                                                    proyecto=ProyectosInvestigacion.objects.get(pk=request.POST['idproyecto']),
                                                    carreras_participantes= CarrerasParticipantes.objects.get(pk=request.POST['idcarrera']),
                                                    resultados_aprendizaje= f.cleaned_data['resultados_aprendizaje'],
                                                    perfil= f.cleaned_data['perfil']
                    )
                    perfil.save(request)

                    perfil.asignatura.clear()
                    for asig in f.cleaned_data['asignatura']:
                        perfil.asignatura.add(asig)

                    request.session['pestaña'] = 'carreraParticipante'
                    request.session['vista'] = '3'
                    log(u'Adiciono perfil al proyecto: %s' % perfil, request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                     raise()
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'editperfilprofesional':
            try:
                f = PerfilProfesionalForm(request.POST)
                if f.is_valid():
                    perfil = PerfilProfesional.objects.get(pk = request.POST['id'])
                    perfil.resultados_aprendizaje= f.cleaned_data['resultados_aprendizaje']
                    perfil.perfil = f.cleaned_data['perfil']
                    perfil.save(request)

                    perfil.asignatura.clear()
                    for asig in f.cleaned_data['asignatura']:
                        perfil.asignatura.add(asig)

                    request.session['pestaña'] = 'carreraParticipante'
                    request.session['vista'] = '3'
                    log(u'Edito perfil al proyecto: %s' % perfil, request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                     raise()
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'addbeneficiario':
            try:
                f = BeneficiariosForm(request.POST)
                if f.is_valid():
                    beneficiario = Beneficiarios(
                        proyecto=ProyectosInvestigacion.objects.get(pk=request.POST['idproyecto']),
                        nombre=f.cleaned_data['nombre'],
                        direccion=f.cleaned_data['direccion'],
                        representante=f.cleaned_data['representante'],
                        cargo_repre=f.cleaned_data['cargo_repre'],
                        telefono=f.cleaned_data['telefono'],
                        correo=f.cleaned_data['correo'],
                        coordenadas=f.cleaned_data['coordenadas'],
                        num_beneficiario_directo=f.cleaned_data['num_beneficiario_directo'],
                        num_beneficiario_indirecto=f.cleaned_data['num_beneficiario_indirecto'],
                        especifico=f.cleaned_data['especifico'],
                        caracteristica=f.cleaned_data['caracteristica'],
                    )
                    beneficiario.save(request)
                    if 'archivo' in request.FILES:
                        newfile = request.FILES['archivo']
                        newfile._name = generar_nombre("proyecto_", newfile._name)

                        beneficiario.archivo = newfile
                        beneficiario.save(request)
                    request.session['pestaña'] = 'beneficiario'
                    request.session['vista'] = '4'
                    log(u'Adiciono beneficiario al proyecto: %s' % beneficiario, request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                     raise()
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'editbeneficiario':
            try:
                f = BeneficiariosForm(request.POST)
                if f.is_valid():
                    beneficiario = Beneficiarios.objects.get(pk = request.POST['id'])
                    beneficiario.nombre = f.cleaned_data['nombre']
                    beneficiario.direccion = f.cleaned_data['direccion']
                    beneficiario.representante = f.cleaned_data['representante']
                    beneficiario.cargo_repre = f.cleaned_data['cargo_repre']
                    beneficiario.telefono = f.cleaned_data['telefono']
                    beneficiario.correo = f.cleaned_data['correo']
                    beneficiario.coordenadas = f.cleaned_data['coordenadas']
                    beneficiario.num_beneficiario_directo = f.cleaned_data['num_beneficiario_directo']
                    beneficiario.num_beneficiario_indirecto = f.cleaned_data['num_beneficiario_indirecto']
                    beneficiario.especifico = f.cleaned_data['especifico']
                    beneficiario.caracteristica = f.cleaned_data['caracteristica']

                    beneficiario.save(request)
                    if 'archivo' in request.FILES:
                        newfile = request.FILES['archivo']
                        newfile._name = generar_nombre("proyecto_", newfile._name)

                        beneficiario.archivo = newfile
                        beneficiario.save(request)
                    request.session['pestaña'] = 'beneficiario'
                    request.session['vista'] = '4'
                    log(u'Edito beneficiario al proyecto: %s' % beneficiario, request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                     raise()
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'deletebeneficiario':
            try:
                participante = Beneficiarios.objects.get(pk=request.POST['id'])
                participante.status = False
                log(u'Elimino beneficiario del proyecto: %s' % participante, request, "del")
                participante.save(request)
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        elif action == 'addinvolucrado':
            try:
                f = InvolucradoForm(request.POST)
                if f.is_valid():
                    involucrado = Involucrado(
                        proyecto=ProyectosInvestigacion.objects.get(pk=request.POST['idproyecto']),
                        nombre=f.cleaned_data['nombre'],
                        direccion=f.cleaned_data['direccion'],
                        representante=f.cleaned_data['representante'],
                        cargo_repre=f.cleaned_data['cargo_repre'],
                        telefono=f.cleaned_data['telefono'],
                        correo=f.cleaned_data['correo'],
                        coordenadas=f.cleaned_data['coordenadas'],
                        tipoActor=f.cleaned_data['tipoActor'],
                    )
                    involucrado.save(request)
                    if 'archivo' in request.FILES:
                        newfile = request.FILES['archivo']
                        newfile._name = generar_nombre("proyecto_", newfile._name)

                        involucrado.archivo = newfile
                        involucrado.save(request)
                    request.session['pestaña'] = 'beneficiario'
                    request.session['vista'] = '4'
                    log(u'Adiciono involucrado al proyecto: %s' % involucrado, request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                     raise()
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'editinvolucrado':
            try:
                f = InvolucradoForm(request.POST)
                if f.is_valid():
                    involucrado = Involucrado.objects.get(pk = request.POST['id'])
                    involucrado.nombre = f.cleaned_data['nombre']
                    involucrado.direccion = f.cleaned_data['direccion']
                    involucrado.representante = f.cleaned_data['representante']
                    involucrado.cargo_repre = f.cleaned_data['cargo_repre']
                    involucrado.telefono = f.cleaned_data['telefono']
                    involucrado.correo = f.cleaned_data['correo']
                    involucrado.coordenadas = f.cleaned_data['coordenadas']
                    involucrado.tipoActor = f.cleaned_data['tipoActor']
                    involucrado.save(request)
                    if 'archivo' in request.FILES:
                        newfile = request.FILES['archivo']
                        newfile._name = generar_nombre("proyecto_", newfile._name)

                        involucrado.archivo = newfile
                        involucrado.save(request)
                    request.session['pestaña'] = 'beneficiario'
                    request.session['vista'] = '4'
                    log(u'Edito involucrado al proyecto: %s' % involucrado, request, "edi")
                    return JsonResponse({"result": "ok"})
                else:
                     raise()
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'deleteinvolucrado':
            try:
                involucrado = Involucrado.objects.get(pk=request.POST['id'])
                involucrado.status = False
                involucrado.save(request)
                log(u'Elimino involucrado del proyecto: %s' % involucrado, request, "del")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        elif action == 'editArbolProblema':
            try:
                f = ArProblemaForm(request.POST)
                if f.is_valid():
                    # Arbol Problema
                    datos = ArbolProblema.objects.get(pk=(request.POST['id']))
                    datos.detalle = f.cleaned_data['detalle']
                    datos.editado = True
                    datos.save(request)
                    # Linea Base
                    datosLinBase = MatrizLineaBase.objects.get(arbolProblema_id=request.POST['id'])
                    datosLinBase.detalle = f.cleaned_data['detalle']
                    datosLinBase.editado = False
                    datosLinBase.save(request)
                    # Arbol Objetivo
                    datosArObj = ArbolObjetivo.objects.get(arbolProblema_id=request.POST['id'])
                    datosArObj.detalle = f.cleaned_data['detalle']
                    datosArObj.editado = False
                    datosArObj.save(request)
                    # Marco Logico
                    if MarcoLogico.objects.filter(arbolObjetivo_id=datosArObj.id).exists():
                        datosMarLogico = MarcoLogico.objects.get(arbolObjetivo_id=datosArObj.id)
                        datosMarLogico.resumen_narrativo = f.cleaned_data['detalle']
                        datosMarLogico.editado = False
                        datosMarLogico.save(request)

                    log(u'Edito Arbol de problema al proyecto: %s' % datos, request, "edit")
                    request.session['pestaña'] = 'aproblema'
                    request.session['vista'] = '5'
                    return JsonResponse({"result": "ok"})
                else:
                     raise()
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'addSubCausa':
            try:
                f = ArProblemaForm(request.POST)
                if f.is_valid():
                    auxCausa = ArbolProblema.objects.get(pk=(request.POST['id']))
                    auxSubCausa = len(ArbolProblema.objects.filter(status=True, parentID=auxCausa, tipo=2))
                    causa = ArbolProblema(
                        proyecto=ProyectosInvestigacion.objects.get(pk=(request.POST['pry'])),
                        parentID=ArbolProblema.objects.get(pk=(request.POST['id'])),
                        orden=auxCausa.orden + '.' + str(auxSubCausa + 1),
                        tipo=2,
                        editado=True,
                        detalle=f.cleaned_data['detalle'],
                    )
                    causa.save(request)
                    medio = ArbolObjetivo(
                        proyecto=ProyectosInvestigacion.objects.get(pk=(request.POST['pry'])),
                        arbolProblema=causa,
                        parentID=ArbolObjetivo.objects.get(arbolProblema_id=(request.POST['id'])),
                        orden=auxCausa.orden + '.' + str(auxSubCausa + 1),
                        tipo=2,
                        detalle=f.cleaned_data['detalle'],
                    )
                    medio.save(request)
                    causalinBase = MatrizLineaBase(
                        proyecto=ProyectosInvestigacion.objects.get(pk=(request.POST['pry'])),
                        arbolProblema=causa,
                        descripcion=f.cleaned_data['detalle'],
                    )
                    causalinBase.save(request)
                    acciones = MarcoLogico(
                        proyecto=ProyectosInvestigacion.objects.get(pk=(request.POST['pry'])),
                        arbolObjetivo=medio,
                        resumen_narrativo=f.cleaned_data['detalle'],
                    )
                    acciones.save(request)

                    log(u'Adiciono subcausa Arbol de problema al proyecto: %s' % auxCausa, request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                     raise()
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'addCausaEfecto':
            try:
                f = ArProb_CausaEfectoForm(request.POST)
                if f.is_valid():
                    auxCausa = len(
                        ArbolProblema.objects.filter(proyecto=request.POST['pry'], status=True,parentID=None, tipo=2))
                    auxEfecto = len(
                        ArbolProblema.objects.filter(proyecto=request.POST['pry'], status=True,parentID=None, tipo=3))
                    # Arbol Problema
                    causa = ArbolProblema(
                        proyecto=ProyectosInvestigacion.objects.get(pk=request.POST['pry']),
                        detalle=f.cleaned_data['causa'],
                        tipo=2,
                        orden=auxCausa + 1,
                        editado=True,
                    )
                    causa.save(request)
                    efecto = ArbolProblema(
                        proyecto=ProyectosInvestigacion.objects.get(pk=request.POST['pry']),
                        detalle=f.cleaned_data['efecto'],
                        tipo=3,
                        orden=auxCausa + 1,
                        editado=True,
                    )
                    efecto.save(request)
                    # Linea base
                    causalinBase = MatrizLineaBase(
                        proyecto=ProyectosInvestigacion.objects.get(pk=request.POST['pry']),
                        arbolProblema=causa,
                        # descripcion=f.cleaned_data['causa'],
                    )
                    causalinBase.save(request)
                    efectolinBase = MatrizLineaBase(
                        proyecto=ProyectosInvestigacion.objects.get(pk=request.POST['pry']),
                        arbolProblema=efecto,
                        descripcion=f.cleaned_data['efecto'],
                    )
                    efectolinBase.save(request)
                    # Arbol Objetivo
                    medio = ArbolObjetivo(
                        proyecto=ProyectosInvestigacion.objects.get(pk=request.POST['pry']),
                        arbolProblema=causa,
                        detalle=f.cleaned_data['causa'],
                        tipo=2,
                        orden=auxCausa + 1,
                    )
                    medio.save(request)
                    fin = ArbolObjetivo(
                        proyecto=ProyectosInvestigacion.objects.get(pk=request.POST['pry']),
                        arbolProblema=efecto,
                        detalle=f.cleaned_data['efecto'],
                        tipo=3,
                        orden=auxEfecto + 1,
                    )
                    fin.save(request)
                    # Marco Logico
                    componente = MarcoLogico(
                        proyecto=ProyectosInvestigacion.objects.get(pk=request.POST['pry']),
                        arbolObjetivo=medio,
                        resumen_narrativo=f.cleaned_data['causa'],
                        editado=True,
                    )
                    componente.save(request)
                    # fin
                    finalidad = MarcoLogico(
                        proyecto=ProyectosInvestigacion.objects.get(pk=request.POST['pry']),
                        arbolObjetivo=ArbolObjetivo.objects.get(pk=fin.pk),
                        resumen_narrativo=fin.detalle,
                    )
                    finalidad.save(request)

                    # subcausas
                    auxSubCausa = len(ArbolProblema.objects.filter(status=True, parentID=causa, tipo=2))
                    subcausa = ArbolProblema(
                        proyecto=ProyectosInvestigacion.objects.get(pk=(request.POST['pry'])),
                        parentID=causa,
                        orden=str(causa.orden) + '.' + str(auxSubCausa + 1),
                        tipo=2,
                        detalle=f.cleaned_data['causa'],
                    )
                    subcausa.save(request)
                    medio = ArbolObjetivo(
                        proyecto=ProyectosInvestigacion.objects.get(pk=(request.POST['pry'])),
                        arbolProblema=subcausa,
                        parentID=ArbolObjetivo.objects.get(arbolProblema=causa),
                        orden=str(causa.orden) + '.' + str(auxSubCausa + 1),
                        tipo=2,
                        detalle=f.cleaned_data['causa'],
                    )
                    medio.save(request)
                    causalinBase = MatrizLineaBase(
                        proyecto=ProyectosInvestigacion.objects.get(pk=(request.POST['pry'])),
                        arbolProblema=subcausa,
                        descripcion=f.cleaned_data['causa'],
                    )
                    causalinBase.save(request)
                    acciones = MarcoLogico(
                        proyecto=ProyectosInvestigacion.objects.get(pk=(request.POST['pry'])),
                        arbolObjetivo=medio,
                        resumen_narrativo=f.cleaned_data['causa'],
                    )
                    acciones.save(request)

                    log(u'Adiciono causa efecto Arbol de problema al proyecto: %s' % auxCausa, request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                     raise()
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'deleteArbolProblema':
            try:
                total = 0.0
                if ArbolProblema.objects.filter(status=True, parentID_id=request.POST['id']).exists():
                    subelemento = ArbolProblema.objects.filter(status=True, parentID_id=request.POST['id'])
                    for sub in subelemento:
                        arPro = ArbolProblema.objects.get(id=sub.pk)
                        arPro.status = False
                        arPro.save(request)
                        arObj = ArbolObjetivo.objects.get(arbolProblema_id=arPro.id)
                        arObj.status = False
                        arObj.save(request)
                        linBase = MatrizLineaBase.objects.get(arbolProblema_id=arPro.id)
                        linBase.status = False
                        linBase.save(request)
                        if MarcoLogico.objects.filter(arbolObjetivo_id=arObj.id).exists():
                            marLogi = MarcoLogico.objects.get(arbolObjetivo_id=arObj.id)
                            marLogi.status = False
                            marLogi.save(request)
                        if Presupuesto.objects.filter(status=True, aobjetivo=arObj).exists():
                            presupuesto = Presupuesto.objects.filter(status=True, aobjetivo=arObj)
                            for presupuesto in presupuesto:
                                presupuesto.status = False
                                presupuesto.save(request)
                            for pre in Presupuesto.objects.filter(status=True, proyecto=arObj.proyecto):
                                total = total + float(pre.total)
                            pry = ProyectosInvestigacion.objects.get(id=arObj.proyecto.id)
                            pry.valorpresupuestointerno = total
                            pry.save(request)

                if not ArbolProblema.objects.filter(status=True, parentID_id=request.POST['id']).exists():
                    arPro = ArbolProblema.objects.get(id=request.POST['id'])
                    arPro.status = False
                    arPro.save(request)
                    arObj = ArbolObjetivo.objects.get(arbolProblema_id=arPro.id)
                    arObj.status = False
                    arObj.save(request)
                    linBase = MatrizLineaBase.objects.get(arbolProblema_id=arPro.id)
                    linBase.status = False
                    linBase.save(request)
                    if MarcoLogico.objects.filter(arbolObjetivo_id=arObj.id).exists():
                        marLogi = MarcoLogico.objects.get(arbolObjetivo_id=arObj.id)
                        marLogi.status = False
                        marLogi.save(request)
                    if Presupuesto.objects.filter(status=True, aobjetivo=arObj).exists():
                        presupuesto = Presupuesto.objects.filter(status=True, aobjetivo=arObj)
                        for presupuesto in presupuesto:
                            presupuesto.status = False
                            presupuesto.save(request)
                        for pre in Presupuesto.objects.filter(status=True, proyecto=arObj.proyecto):
                            total = total + float(pre.total)
                        pry = ProyectosInvestigacion.objects.get(id=arObj.proyecto.id)
                        pry.valorpresupuestointerno = total
                        pry.save(request)

                    log(u'Elimino arbol de problema del proyecto: %s' % arPro, request, "del")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        elif action == 'datoLinea':
            try:
                dataRecibida = request.POST['val']
                proyecto = ProyectosInvestigacion.objects.get(pk=request.POST['pry'])
                if dataRecibida == 'true':
                    proyecto.linea_base = True
                else:
                    proyecto.linea_base = False
                proyecto.save(request)
                log(u'edito dato de linea base: %s' % proyecto, request, "edit")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        elif action == 'addDatoSecundario':
            try:
                f = DatoSecundarioForm(request.POST)
                if f.is_valid():
                    datoSecundario = DatoSecundario(
                        proyecto=ProyectosInvestigacion.objects.get(pk=request.POST['pry']),
                        descripcion=f.cleaned_data['descripcion'],
                    )
                    datoSecundario.save(request)
                    request.session['pestaña'] = 'lineaBase'
                    request.session['vista'] = '6'
                    log(u'Adiciono dato secundario al proyecto: %s' % datoSecundario, request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                     raise()
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'editDatoSecundario':
            try:
                f = DatoSecundarioForm(request.POST)
                if f.is_valid():
                    datos = DatoSecundario.objects.get(pk=request.POST['id'])
                    datos.descripcion = f.cleaned_data['descripcion']
                    datos.save(request)
                    request.session['pestaña'] = 'lineaBase'
                    request.session['vista'] = '6'
                    log(u'Edito dato secundario al proyecto: %s' % datos, request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                    raise ()
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'deleteDatoSecundario':
            try:
                datos = DatoSecundario.objects.get(id=request.POST['id'])
                datos.status = False
                datos.save(request)
                log(u'Elimino datos secundarios del proyecto: %s' % datos, request, "del")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        elif action == 'editLineaBase':
            try:
                f = LineaBaseForm(request.POST)
                if f.is_valid():
                    datos = MatrizLineaBase.objects.get(pk=request.POST['id'])
                    datos.descripcion = f.cleaned_data['descripcion']
                    datos.metodo = f.cleaned_data['metodo']
                    datos.fuente = f.cleaned_data['fuente']
                    datos.datos_linea_base = f.cleaned_data['linea_base']
                    datos.editado = True
                    datos.save(request)
                    request.session['pestaña'] = 'lineaBase'
                    request.session['vista'] = '6'
                    log(u'Edito dato secundario al proyecto: %s' % datos, request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                    raise ()
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'editArbolObjetivo':
            try:
                f = ArObjetivoForm(request.POST)
                if f.is_valid():
                    datos = ArbolObjetivo.objects.get(pk=request.POST['id'])
                    datos.detalle = f.cleaned_data['detalle']
                    datos.editado = True
                    datos.save(request)
                    if MarcoLogico.objects.filter(arbolObjetivo_id=request.POST['id']).exists():
                        datosMarLogico = MarcoLogico.objects.get(arbolObjetivo_id=request.POST['id'])
                        datosMarLogico.resumen_narrativo = f.cleaned_data['detalle']
                        datosMarLogico.editado = False
                        datosMarLogico.save(request)
                    datos.save(request)
                    request.session['pestaña'] = 'arbol'
                    request.session['vista'] = '7'
                    log(u'Edito arbol de objetivo: %s' % datos, request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                    raise ()
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'editMarcoLogico':
            try:
                f = MarcoLogicoForm2(request.POST)
                if f.is_valid():
                    datos = MarcoLogico.objects.get(pk=request.POST['id'])
                    datos.resumen_narrativo = f.cleaned_data['resumen_narrativo']
                    datos.indicador = f.cleaned_data['indicador']
                    datos.fuente = f.cleaned_data['fuente']
                    datos.supuestos = f.cleaned_data['supuestos']
                    datos.cumplimiento = f.cleaned_data['cumplimiento']
                    datos.editado = True
                    datos.save(request)


                    acciones = ArbolObjetivo.objects.filter(parentID = datos.arbolObjetivo, status=True)
                    if acciones.exists():
                        if datos.cumplimiento == "0.00":
                            cumpl_acciones=0.00
                        else:
                            cumpl_acciones = float(datos.cumplimiento)/acciones.count()
                        for acc in acciones:
                            actividad = MarcoLogico.objects.get(arbolObjetivo=acc)
                            actividad.cumplimiento = cumpl_acciones
                            actividad.save(request)

                    request.session['pestaña'] = 'marco'
                    request.session['vista'] = '8'
                    log(u'Edito marco logico: %s' % datos, request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                    raise ()
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'addcronograma':
            try:
                f = CronogramaForm(request.POST)
                if f.is_valid():
                    aobj = ArbolObjetivo.objects.get(pk=request.POST['aObj'])
                    porcent = MarcoLogico.objects.get(arbolObjetivo=aobj).cumplimiento
                    tareas = Cronograma.objects.filter(status=True, aobjetivo = aobj)

                    if (f.cleaned_data['fecha_inicio'] >= aobj.proyecto.fechainicio and f.cleaned_data['fecha_inicio'] <= aobj.proyecto.fechaplaneacion) and (f.cleaned_data['fecha_fin'] >= aobj.proyecto.fechainicio and f.cleaned_data['fecha_fin'] <= aobj.proyecto.fechaplaneacion):
                        cronograma = Cronograma(
                            proyecto=aobj.proyecto,
                            aobjetivo=aobj,
                            descripcion=f.cleaned_data['descripcion'],
                            fecha_inicio=f.cleaned_data['fecha_inicio'],
                            fecha_fin=f.cleaned_data['fecha_fin'],
                        )
                        cronograma.save(request)

                        cronograma.responsable.clear()
                        for resp in f.cleaned_data['responsable']:
                            cronograma.responsable.add(resp)

                        for tarea in tareas:
                            tarea.cumplimiento = porcent/tareas.count()
                            tarea.save(request)
                        request.session['pestaña'] = 'cronograma'
                        request.session['vista'] = '9'
                    else:
                        return JsonResponse({"result": "bad", "mensaje": u"Debe ingresar una fecha dentro del rango de ejecución"})
                        # fechainicio__lte = datetime.now(), fechafin__gte = datetime.now()

                    log(u'Adiciono cronograma al proyecto: %s' % cronograma, request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                     raise()
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'editcronograma':
            try:
                f = CronogramaForm(request.POST)
                if f.is_valid():
                    datos = Cronograma.objects.get(pk=request.POST['id'])
                    if (f.cleaned_data['fecha_inicio'] >= datos.proyecto.fechainicio and f.cleaned_data['fecha_inicio'] <= datos.proyecto.fechaplaneacion) and (f.cleaned_data['fecha_fin'] >= datos.proyecto.fechainicio and f.cleaned_data['fecha_fin'] <= datos.proyecto.fechaplaneacion):

                        datos.descripcion = f.cleaned_data['descripcion']
                        datos.fecha_inicio = f.cleaned_data['fecha_inicio']
                        datos.fecha_fin = f.cleaned_data['fecha_fin']
                        datos.save(request)

                        datos.responsable.clear()
                        for resp in f.cleaned_data['responsable']:
                            datos.responsable.add(resp)
                        request.session['pestaña'] = 'cronograma'
                        request.session['vista'] = '9'
                    else:
                        return JsonResponse({"result": "bad", "mensaje": u"Debe ingresar una fecha dentro del rango de ejecución"})
                    log(u'edito cronograma al proyecto: %s' % datos, request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                     raise()
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'deletecronograma':
            try:
                datos = Cronograma.objects.get(id=request.POST['id'])
                aobj = ArbolObjetivo.objects.get(pk=datos.aobjetivo.pk)
                datos.status = False
                datos.save(request)

                porcent = MarcoLogico.objects.get(arbolObjetivo=aobj).cumplimiento
                tareas = Cronograma.objects.filter(status=True, aobjetivo=aobj)

                for tarea in tareas:
                    tarea.cumplimiento = porcent / tareas.count()
                    tarea.save(request)
                request.session['pestaña'] = 'cronograma'
                request.session['vista'] = '9'
                log(u'Elimino cronograma del proyecto: %s' % datos, request, "del")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        elif action == 'addpresupuesto':
            try:
                f = PresupuestoForm(request.POST)
                if f.is_valid():
                    aobj = ArbolObjetivo.objects.get(pk=request.POST['aObj'])
                    pry = ProyectosInvestigacion.objects.get(pk=aobj.proyecto.pk)
                    presupuesto = Presupuesto(
                        proyecto=aobj.proyecto,
                        aobjetivo=aobj,
                        cantidad=f.cleaned_data['cantidad'],
                        suministro=f.cleaned_data['suministro'],
                        costo_unitario=f.cleaned_data['costo_unitario'],
                        subtotal=f.cleaned_data['subtotal'],
                        iva=f.cleaned_data['iva'],
                        total=f.cleaned_data['total'],
                        especificaciones=f.cleaned_data['especificaciones'],
                    )
                    presupuesto.save(request)
                    total = 0.0
                    for pre in Presupuesto.objects.filter(status=True, proyecto=pry):
                        total = total + float(pre.total)
                    pry.valorpresupuestointerno = total
                    pry.save(request)
                    request.session['pestaña'] = 'presupuesto'
                    request.session['vista'] = '10'
                    log(u'Adiciono presupuesto al proyecto: %s' % presupuesto, request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                     raise()
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'editpresupuesto':
            try:
                f = PresupuestoForm(request.POST)
                if f.is_valid():
                    datos = Presupuesto.objects.get(pk=request.POST['id'])
                    pry = ProyectosInvestigacion.objects.get(pk=datos.proyecto.pk)
                    datos.cantidad = f.cleaned_data['cantidad']
                    datos.producto = None
                    datos.suministro = f.cleaned_data['suministro']
                    datos.insumo = None
                    datos.costo_unitario = f.cleaned_data['costo_unitario']
                    datos.subtotal = f.cleaned_data['subtotal']
                    datos.iva = f.cleaned_data['iva']
                    datos.total = f.cleaned_data['total']
                    datos.otros = False
                    datos.especificaciones = f.cleaned_data['especificaciones']
                    datos.save(request)

                    total = 0.0
                    for pre in Presupuesto.objects.filter(status=True, proyecto=pry):
                        total = total + float(pre.total)
                    pry.valorpresupuestointerno = total
                    pry.save(request)

                    request.session['pestaña'] = 'presupuesto'
                    request.session['vista'] = '10'
                    log(u'Edito presupuesto al proyecto: %s' % datos, request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                     raise()
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'deletepresupuesto':
            try:
                datos = Presupuesto.objects.get(id=request.POST['id'])
                datos.status = False
                datos.save(request)
                log(u'Elimino presupuesto del proyecto: %s' % datos, request, "del")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        elif action == 'addredaccion':
            try:
                f = RedaccionForm(request.POST)
                if f.is_valid():
                    redaccion = Problema(
                        proyecto=ProyectosInvestigacion.objects.get(pk=request.POST['pry']),
                        descripcion=f.cleaned_data['descripcion'],
                        antecedentes=f.cleaned_data['antecedentes'],
                        deteccion_necesidades=f.cleaned_data['deteccion_necesidades'],
                        justificacion=f.cleaned_data['justificacion'],
                        Pertinencia=f.cleaned_data['Pertinencia'],
                        metodologia=f.cleaned_data['metodologia'],
                        seguimiento=f.cleaned_data['seguimiento'],
                        evaluacion=f.cleaned_data['evaluacion'],
                        producto=f.cleaned_data['producto'],
                        Bibliografia=f.cleaned_data['Bibliografia'],
                        objetivo_general=f.cleaned_data['objetivo_general'],
                        objetivos_especificos=f.cleaned_data['objetivos_especificos'],
                        convenio=f.cleaned_data['convenio'],
                    )
                    redaccion.save(request)
                    request.session['pestaña'] = 'redaccion'
                    request.session['vista'] = '11'
                    log(u'Adiciono redaccion al proyecto: %s' % redaccion, request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                     raise()
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'editredaccion':
            try:
                f = RedaccionForm(request.POST)
                if f.is_valid():
                    datos = Problema.objects.get(pk=request.POST['id'])
                    datos.descripcion = f.cleaned_data['descripcion']
                    datos.antecedentes = f.cleaned_data['antecedentes']
                    datos.deteccion_necesidades = f.cleaned_data['deteccion_necesidades']
                    datos.justificacion = f.cleaned_data['justificacion']
                    datos.Pertinencia = f.cleaned_data['Pertinencia']
                    datos.metodologia = f.cleaned_data['metodologia']
                    datos.seguimiento = f.cleaned_data['seguimiento']
                    datos.evaluacion = f.cleaned_data['evaluacion']
                    datos.producto = f.cleaned_data['producto']
                    datos.Bibliografia = f.cleaned_data['Bibliografia']
                    datos.objetivo_general = f.cleaned_data['objetivo_general']
                    datos.objetivos_especificos = f.cleaned_data['objetivos_especificos']
                    datos.convenio = f.cleaned_data['convenio']
                    datos.save(request)
                    request.session['pestaña'] = 'redaccion'
                    request.session['vista'] = '11'
                    log(u'edito redaccion al proyecto: %s' % datos, request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                     raise()
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'addanexo':
            try:
                f = AnexosForm(request.POST)
                if f.is_valid():
                    anexo = Anexos(
                        proyecto=ProyectosInvestigacion.objects.get(pk=request.POST['pry']),
                        titulo=f.cleaned_data['titulo'],
                    )
                    anexo.save(request)
                    if 'archivo' in request.FILES:
                        newfile = request.FILES['archivo']
                        newfile._name = generar_nombre("proyecto_", newfile._name)

                        anexo.archivo = newfile
                        anexo.save(request)

                    log(u'Adiciono anexo al proyecto: %s' % anexo, request, "add")
                    request.session['pestaña'] = 'anexo'
                    request.session['vista'] = '12'
                    return JsonResponse({"result": "ok"})
                else:
                     raise()
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'editanexo':
            try:
                f = AnexosForm(request.POST)
                if f.is_valid():
                    anexo = Anexos.objects.get(pk=request.POST['id'])
                    anexo.titulo = f.cleaned_data['titulo']
                    anexo.save(request)
                    if 'archivo' in request.FILES:
                        newfile = request.FILES['archivo']
                        newfile._name = generar_nombre("proyecto_", newfile._name)

                        anexo.archivo = newfile
                        anexo.save(request)

                    log(u'Edito anexo al proyecto: %s' % anexo, request, "edit")
                    request.session['pestaña'] = 'anexo'
                    request.session['vista'] = '12'
                    return JsonResponse({"result": "ok"})
                else:
                     raise()
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'deleteanexo':
            try:
                datos = Anexos.objects.get(id=request.POST['id'])
                datos.status = False
                datos.save(request)
                log(u'Elimino anexo del proyecto: %s' % datos, request, "del")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        elif action == 'addsuministro':
            try:
                f = SuministroForm(request.POST)
                if f.is_valid():
                    suministro = Suministro(
                        rubro=f.cleaned_data['rubro'],
                        especificacion=f.cleaned_data['especificacion'],
                        costo_unitario=f.cleaned_data['costo_unitario'],
                        aplicaIva=f.cleaned_data['aplica_iva'],
                        activo=f.cleaned_data['activo'],
                    )
                    suministro.save(request)
                    log(u'Adicionó suministro para proyecto: %s' % suministro, request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                     raise()
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'editsuministro':
            try:
                f = SuministroForm(request.POST)
                if f.is_valid():
                    suministro = Suministro.objects.get(pk=request.POST['id'])
                    suministro.rubro = f.cleaned_data['rubro']
                    suministro.especificacion = f.cleaned_data['especificacion']
                    suministro.costo_unitario = f.cleaned_data['costo_unitario']
                    suministro.aplicaIva = f.cleaned_data['aplica_iva']
                    suministro.activo = f.cleaned_data['activo']
                    suministro.save(request)

                    log(u'Editó suministro para proyecto: %s' % suministro, request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                     raise()
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'deletesuministro':
            try:
                datos = Suministro.objects.get(id=request.POST['id'])
                datos.status = False
                datos.save(request)
                log(u'Elimino suministro para proyecto: %s' % datos, request, "del")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        elif action == 'evaluarinformes':
            try:
                informe = InformesProyectoVinculacionEstudiante.objects.get(pk=int(request.POST['id']))
                informe.estado = request.POST['est']
                informe.observacion = request.POST['obs']
                informe.fecharevisionarchivo = datetime.now().date()
                informe.save(request)

                asunto = "Revisión de informes de Proyecto de Vinculación."
                mensaje = "ha sido revisado su informe"
                template = "emails/notificarrevisioninformevinculacion.html"
                datos = {'sistema': u'SGA - UNEMI',
                         'mensaje': mensaje,
                         'fecha': datetime.now().date(),
                         'hora': datetime.now().time(),
                         'saludo': 'Estimada' if informe.proyecto.inscripcion.persona.sexo_id == 1 else 'Estimado',
                         'estudiante': informe.proyecto.inscripcion.persona.nombre_completo_inverso(),
                         'estado': request.POST['est'],
                         'informe': informe,
                         'proyecto': informe.proyecto.proyecto.nombre,
                         'autoridad2': '',
                         't': miinstitucion()
                         }
                email = informe.proyecto.inscripcion.persona.lista_emails_envio()

                send_html_mail(asunto, template, datos, email, [], cuenta=variable_valor('CUENTAS_CORREOS')[0])

                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'solicitarevision':
            try:
                datos = ProyectosInvestigacion.objects.get(id=request.POST['id'])
                if FechaProyectos.objects.values('id').filter(status=True, fechainicio__lte=datetime.now(),fechafin__gte=datetime.now()).exists():
                    conv = FechaProyectos.objects.filter(status=True, fechainicio__lte=datetime.now(), fechafin__gte=datetime.now()).last()
                    datos.convocatoria = conv
                    datos.aprobacion = 4
                    datos.save(request)
                    log(u'Solicito revision de proyecto: %s' % datos, request, "edit")
                else:
                    return JsonResponse({"result": "bad", "mensaje": u"Actualmente no existe un periodo de convocatoria activa para solicitar revisión"})
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        elif action == 'addfechafinp':
            try:

                f = FechaFinProyectoFrom(request.POST)
                proyecto = ProyectosInvestigacion.objects.get(pk=request.POST['id'])
                if f.is_valid():
                    proyecto.fechareal = f.cleaned_data['fechafin']
                    proyecto.save(request)
                    log(u'Añadió fecha fin: %s' % proyecto.nombre, request, "add")
                    return JsonResponse({"result": "ok"})

                else:

                    raise NameError('Error')

            except Exception as ex:

                transaction.set_rollback(True)

                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'editprograma':
            try:
                f = ProgramasVinculacionForm(request.POST)
                programas = ProgramasInvestigacion.objects.get(pk=request.POST['id'])
                if f.is_valid():
                    programas.fechainicio = f.cleaned_data['fechainicio']
                    programas.fechaplaneado = f.cleaned_data['fechaplaneado']
                    programas.fechareal = f.cleaned_data['fechareal']
                    programas.nombre = f.cleaned_data['nombre']
                    # programas.lineainvestigacion = f.cleaned_data['lineainvestigacion']
                    programas.alcanceterritorial = f.cleaned_data['alcanceterritorial']
                    # programas.areaconocimiento = f.cleaned_data['areaconocimiento']
                    # programas.subareaconocimiento = f.cleaned_data['subareaconocimiento']
                    # programas.subareaespecificaconocimiento = f.cleaned_data['subareaespecificaconocimiento']
                    programas.tiempoejecucion = f.cleaned_data['tiempoejecucion']
                    # programas.sublineainvestigacion = f.cleaned_data['sublineainvestigacion']
                    programas.valorpresupuestointerno = f.cleaned_data['valorpresupuestointerno']
                    programas.valorpresupuestoexterno = f.cleaned_data['valorpresupuestoexterno']
                    programas.save(request)
                    if programas.programasvinculacioncampos_set.filter(status=True).exists():
                        programascampos = programas.programasvinculacioncampos_set.filter(status=True)[0]
                        programascampos.soportedelprograma = f.cleaned_data['soportedelprograma']
                        programascampos.elaboradopor = f.cleaned_data['elaboradopor']
                        programascampos.perfilbeneficiarios = f.cleaned_data['perfilbeneficiarios']
                        programascampos.numerobeneficiarios = f.cleaned_data['numerobeneficiarios']
                        programascampos.proyectosintegranprograma = f.cleaned_data['proyectosintegranprograma']
                        programascampos.perfildocentesestudiantes = f.cleaned_data['perfildocentesestudiantes']
                        programascampos.planteamientoproblema = f.cleaned_data['planteamientoproblema']
                        programascampos.justificacion = f.cleaned_data['justificacion']
                        programascampos.pertinencia = f.cleaned_data['pertinencia']
                        programascampos.objetivogeneral = f.cleaned_data['objetivogeneral']
                        programascampos.objetivoespecifico = f.cleaned_data['objetivoespecifico']
                        programascampos.metodologia = f.cleaned_data['metodologia']
                        programascampos.recursos = f.cleaned_data['recursos']
                        programascampos.cronograma = f.cleaned_data['cronograma']
                        programascampos.seguimiento = f.cleaned_data['seguimiento']
                        programascampos.evaluacion = f.cleaned_data['evaluacion']
                        programascampos.bibliografia = f.cleaned_data['bibliografia']
                        programascampos.anexos = f.cleaned_data['anexos']
                    else:
                        programascampos = ProgramasVinculacionCampos(programasvinculacion=programas,
                                                                    soportedelprograma=f.cleaned_data['soportedelprograma'],
                                                                    elaboradopor=f.cleaned_data['elaboradopor'],
                                                                    perfilbeneficiarios=f.cleaned_data['perfilbeneficiarios'],
                                                                    numerobeneficiarios=f.cleaned_data['numerobeneficiarios'],
                                                                    proyectosintegranprograma=f.cleaned_data['proyectosintegranprograma'],
                                                                    perfildocentesestudiantes=f.cleaned_data['perfildocentesestudiantes'],
                                                                    planteamientoproblema=f.cleaned_data['planteamientoproblema'],
                                                                    justificacion=f.cleaned_data['justificacion'],
                                                                    pertinencia=f.cleaned_data['pertinencia'],
                                                                    objetivogeneral=f.cleaned_data['objetivogeneral'],
                                                                    objetivoespecifico=f.cleaned_data['objetivoespecifico'],
                                                                    metodologia=f.cleaned_data['metodologia'],
                                                                    recursos=f.cleaned_data['recursos'],
                                                                    cronograma=f.cleaned_data['cronograma'],
                                                                    seguimiento=f.cleaned_data['seguimiento'],
                                                                    evaluacion=f.cleaned_data['evaluacion'],
                                                                    bibliografia=f.cleaned_data['bibliografia'],
                                                                    anexos=f.cleaned_data['anexos'])

                    programascampos.save(request)
                    log(u'Editó programa de investigación: %s' % programas, request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                     raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'editaprobacion':
            try:
                f = AprobacionProyecto(request.POST)
                periodoarea = ProyectosInvestigacion.objects.get(pk=request.POST['id'])

                if f.is_valid():
                    newfile = None
                    if 'archivo' in request.FILES:
                        newfile = request.FILES['archivo']

                    periodoarea.aprobacion = f.cleaned_data['area']
                    periodoarea.observaa = f.cleaned_data['nombre'].upper()
                    periodoarea.save(request)

                    if not f.cleaned_data['area']==1:

                        aprobacion = ProyectosInvestigacionAprobacion(
                            proyecto = periodoarea,
                            observacion = f.cleaned_data['nombre'].upper(),
                            archivo =newfile
                        )
                        aprobacion.save(request)
                    log(u'edita area a periodo: %s' % periodoarea, request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'addevidenciasprogramas':
            try:
                f = EvidenciaForm(request.POST, request.FILES)
                if f.is_valid():
                    newfile = None
                    if 'archivo' in request.FILES:
                        newfile = request.FILES['archivo']
                        newfile._name = generar_nombre("programa_", newfile._name)
                    if DetalleEvidencias.objects.filter(evidencia_id=request.POST['idevidencia'], programa_id=request.POST['id']).exists():
                        detalle = DetalleEvidencias.objects.get(evidencia_id=request.POST['idevidencia'], programa_id=request.POST['id'])
                        detalle.descripcion = f.cleaned_data['descripcion']
                        detalle.archivo = newfile
                        detalle.save(request)
                        log(u'Editó detalle de evidencia de investigación: %s' % detalle, request, "edit")
                    else:
                        evidencia = DetalleEvidencias(evidencia_id=request.POST['idevidencia'],
                                                      programa_id=request.POST['id'],
                                                      descripcion=f.cleaned_data['descripcion'],
                                                      archivo=newfile)
                        evidencia.save(request)
                        log(u'Adiciono evidencia de investigación: %s' % evidencia, request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                     raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'deleteprograma':
            try:
                programa = ProgramasInvestigacion.objects.get(pk=request.POST['id'])
                programa.status = False
                log(u'Elimino programa de investigación: %s' % programa, request, "del")
                programa.save(request)
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        elif action == 'deleteparticipanteproyecto':
            try:
                # participante = ParticipantesMatrices.objects.get(pk=request.POST['id'])
                # if participante.inscripcion:
                #     participante1= ProyectoVinculacionInscripcion.objects.get(status=True,inscripcion=participante.inscripcion,proyectovinculacion=participante.proyecto)
                #     participante1.status = False
                #     participante1.save(request)
                # participante.status = False
                # log(u'Elimino practicipante de investigación: %s' % participante, request, "del")
                # participante.save(request)
                participante = ParticipantesMatrices.objects.get(pk=request.POST['id'])
                if participante.inscripcion:
                    if ProyectoVinculacionInscripcion.objects.filter(status=True, inscripcion=participante.inscripcion, proyectovinculacion=participante.proyecto).exists():
                        participante1 = ProyectoVinculacionInscripcion.objects.get(status=True, inscripcion=participante.inscripcion, proyectovinculacion=participante.proyecto)
                        participante1.status = False
                        participante1.save(request)
                participante.status = False
                log(u'Elimino practicipante de investigación: %s' % participante, request, "del")
                participante.save(request)
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        elif action == 'deletepresupuestoproyecto':
            try:
                presupuesto = PresupuestosProyecto.objects.get(pk=request.POST['id'])
                presupuesto.status = False
                log(u'Elimino presupuesto de proyecto de investigación: %s' % presupuesto, request, "del")
                presupuesto.save(request)
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        elif action == 'addevidenciasproyectos':
            try:
                f = EvidenciaForm(request.POST, request.FILES)
                # 2.5MB - 2621440
                # 5MB - 5242880
                # 10MB - 10485760
                # 20MB - 20971520
                # 50MB - 5242880
                # 100MB 104857600
                # 250MB - 214958080
                # 500MB - 429916160
                d = request.FILES['archivo']
                if d.size > 10485760:
                    return JsonResponse({"result": "bad", "mensaje": u"Error, archivo mayor a 10 Mb."})
                if f.is_valid():
                    newfile = None
                    if 'archivo' in request.FILES:
                        newfile = request.FILES['archivo']
                        newfile._name = generar_nombre("proyecto_", newfile._name)
                    if DetalleEvidencias.objects.filter(evidencia_id=request.POST['idevidencia'], proyecto_id=request.POST['id']).exists():
                        detalle = DetalleEvidencias.objects.get(evidencia_id=request.POST['idevidencia'], proyecto_id=request.POST['id'])
                        detalle.descripcion = f.cleaned_data['descripcion']
                        detalle.archivo = newfile
                        detalle.save(request)
                        log(u'Adiciono evidencia proyecto de investigación: %s' % detalle, request, "add")
                    else:
                        evidencia = DetalleEvidencias(evidencia_id=request.POST['idevidencia'],
                                                      proyecto_id=request.POST['id'],
                                                      descripcion=f.cleaned_data['descripcion'],
                                                      archivo=newfile)
                        evidencia.save(request)
                        log(u'Adiciono evidencia proyecto de investigación: %s' % evidencia, request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                     raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'addfecha':
            try:
                f = FechaProyectosFrom(request.POST)
                if f.is_valid():
                    fecha =FechaProyectos(
                        descripcion= f.cleaned_data['descripcion'],
                        fechainicio=f.cleaned_data['fechainicio'],
                        fechafin=f.cleaned_data['fechafin']
                    )
                    fecha.save(request)
                    log(u'Adiciono fecha de convocatoria: %s' % fecha.descripcion, request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'editfecha':
            try:
                f = FechaProyectosFrom(request.POST)
                if f.is_valid():
                    convocatoria = FechaProyectos.objects.get(pk=request.POST['id'])
                    convocatoria.descripcion = f.cleaned_data['descripcion']
                    convocatoria.fechainicio = f.cleaned_data['fechainicio']
                    convocatoria.fechafin = f.cleaned_data['fechafin']
                    convocatoria.save(request)
                    log(u'Editó fecha de convocatoria: %s' % convocatoria.descripcion, request, "edit")
                    return JsonResponse({"result": "ok"})

                else:
                     raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'deletefecha':
            try:
                convocatoria = FechaProyectos.objects.get(pk=request.POST['id'])
                convocatoria.status = False
                convocatoria.save(request)
                log(u'Eliminó fecha de convocatoria: %s' % convocatoria, request, "del")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        elif action == 'addfechaentrega':
            try:
                f = FechaEntregaProyectoFrom(request.POST)
                if f.is_valid():
                    proyecto = ProyectosInvestigacion.objects.get(pk=request.POST['id'])
                    proyecto.fecha_entrega = f.cleaned_data['fecha']
                    proyecto.save(request)
                    log(u'Editó fecha de entrega del proecto: %s' % proyecto, request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                     raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'actualizaConvocatoria':
            try:
                f = CambioConvocatoriaForm(request.POST)
                if f.is_valid():
                    proyecto = ProyectosInvestigacion.objects.get(pk=request.POST['id'])
                    proyecto.convocatoria = f.cleaned_data['convocatoria']
                    proyecto.save(request)
                    log(u'Editó convocatoria del proecto: %s' % proyecto, request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                     raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'addfechaaprobacion':
            try:
                f = FechaEntregaProyectoFrom(request.POST)
                if f.is_valid():
                    proyecto = ProyectosInvestigacion.objects.get(pk=request.POST['id'])
                    proyecto.fecha_aprobacion= f.cleaned_data['fecha']
                    proyecto.save(request)
                    log(u'Editó fecha de aprobacion del proecto: %s' % proyecto, request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                     raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'addresolucion':
            try:
                f = ResolucionForm(request.POST)
                if f.is_valid():
                    newfile=None
                    if 'archivo' in request.FILES:
                        newfile = request.FILES['archivo']
                        newfile._name = generar_nombre("resolucion_convocatoria_vinculacion_", newfile._name)
                    resolucion =Resolucion(
                        convocatoria = FechaProyectos.objects.get(pk=request.POST['conv']),
                        resolucion = f.cleaned_data['resolucion'],
                        archivo=newfile
                    )
                    resolucion.save(request)
                    log(u'Añadio resolución: %s' % resolucion, request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                     raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'editresolucion':
            try:
                f = ResolucionForm(request.POST)
                if f.is_valid():
                    resolucion = Resolucion.objects.get(pk=request.POST['id'])
                    resolucion.resolucion= request.POST['resolucion']
                    if 'archivo' in request.FILES:
                        newfile = request.FILES['archivo']
                        newfile._name = generar_nombre("resolucion_", newfile._name)
                        resolucion.archivo = newfile
                    resolucion.save(request)
                    log(u'Editó resolución: %s' % resolucion, request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                     raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'deleteresolucion':
            try:
                resolucion = Resolucion.objects.get(pk=request.POST['id'])
                resolucion.status = False
                resolucion.save(request)
                log(u'Eliminó resolucion: %s' % resolucion, request, "del")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        elif action == 'addcargobeneficiario':
            try:
                f = CargoBeneficiarioFrom(request.POST)
                if f.is_valid():
                    cargo =CargoBeneficiario(
                        descripcion= request.POST['descripcion'],
                    )
                    cargo.save(request)
                    log(u'Adiciono cargo de beneficiario: %s' % cargo, request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                     raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'editcargobeneficiario':
            try:
                f = CargoBeneficiarioFrom(request.POST)
                if f.is_valid():
                    datos = CargoBeneficiario.objects.get(pk=request.POST['id'])
                    datos.descripcion= request.POST['descripcion']
                    datos.save(request)
                    log(u'Editó cargo de beneficiario: %s' % datos, request, "edit")
                    return JsonResponse({"result": "ok"})

                else:
                     raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'deletecargobeneficiario':
            try:
                cargo = CargoBeneficiario.objects.get(pk=request.POST['id'])
                cargo.status = False
                cargo.save(request)
                log(u'Eliminó cargo de beneficiario: %s' % cargo, request, "del")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        elif action == 'addfechafin':
            try:
                f = FechaProyectosFrom(request.POST)
                fecha = request.POST['fechainicio']
                fechaf = request.POST['fechafin']
                fecha  = fecha.split('-')
                fechaf  = fechaf.split('-')

                fecha = date(int(fecha[2]),int(fecha[1]),int(fecha[0]))
                fechaf = date(int(fechaf[2]),int(fechaf[1]),int(fechaf[0]))


                if f.is_valid():
                    #fechainicio=fechainicio=request.POST['fechainicio']
                    #fechafin = fechafin=request.POST['fechafin']
                    fechaa =FechaProyectos(fechainicio=fecha,fechafin=fechaf)
                    fechaa.save(request)
                    #log(u'Adiciono de fecha para proyecto de vinculacion: %s')
                    return JsonResponse({"result": "ok"})

                else:
                     raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'deleteproyecto':
            try:
                proyecto = ProyectosInvestigacion.objects.get(pk=request.POST['id'],tipo=1)
                proyecto.status = False
                log(u'Eliminó proyecto de investigación: %s' % proyecto, request, "add")
                proyecto.save(request)
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        if action == 'addparticipantesdocentes':
            try:
                f = ParticipanteProfesorForm(request.POST)
                if f.is_valid():
                    if ParticipantesMatrices.objects.filter(proyecto_id=request.POST['id'],profesor_id=f.cleaned_data['profesor'],status=True).exists():
                        return JsonResponse({"result": "r","mensaje": u"Error el docente ya se encuentra inscrito en el proyecto."})
                    programas = ParticipantesMatrices(matrizevidencia_id=2,
                                                      proyecto_id=request.POST['id'],
                                                      profesor_id=f.cleaned_data['profesor'],
                                                      horas=f.cleaned_data['horas'],
                                                      tipoparticipante=f.cleaned_data['tipoparticipante']
                                                      )
                    programas.save(request)
                    log(u'Adiciono Participante Docente: %s' % programas, request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                     raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        if action == 'addparticipantesdocentesp':
            try:
                es_externo = False
                externo = None
                personae = None

                f = ParticipanteProfesorVinculacionForm(request.POST)
                if f.is_valid():
                    # Verificar que no esté repetido el integrante
                    if int(f.cleaned_data['tipo']) == 1:
                        if ParticipantesMatrices.objects.filter(proyecto_id=request.POST['id'], profesor_id=f.cleaned_data['profesor'], status=True).exists():
                            return JsonResponse({"result": "bad", "titulo": "Atención!!!", "mensaje": "El docente ya se encuentra inscrito en el proyecto", "showSwal": "True", "swalType": "warning"})
                    else:
                        # Consulto la persona
                        personae = Persona.objects.get(pk=f.cleaned_data['profesor'])

                        # Verifico si tiene perfil externo
                        if Externo.objects.filter(persona=personae, status=True).exists():
                            es_externo = True
                            externo = Externo.objects.get(persona=personae, status=True)

                            if ParticipantesMatrices.objects.filter(proyecto_id=request.POST['id'], externo_id=externo.id, status=True).exists():
                                return JsonResponse({"result": "bad", "titulo": "Atención!!!", "mensaje": "El docente ya se encuentra inscrito en el proyecto", "showSwal": "True", "swalType": "warning"})

                    # Si no existe externo, se debe crear ese perfil, Docentes externos
                    if int(f.cleaned_data['tipo']) == 2:
                        if es_externo is False:
                            # Guardo externo
                            externo = Externo(
                                persona=personae,
                                nombrecomercial='',
                                institucionlabora='',
                                cargodesempena=''
                            )
                            externo.save(request)

                    # Guardo el docente participante del proyecto
                    docente = ParticipantesMatrices(matrizevidencia_id=2,
                                                    proyecto_id=request.POST['id'],
                                                    profesor_id=f.cleaned_data['profesor'] if int(int(f.cleaned_data['tipo'])) == 1 else None,
                                                    externo_id=externo.id if int(int(f.cleaned_data['tipo'])) == 2 else None,
                                                    horas=f.cleaned_data['horas'] if int(int(f.cleaned_data['tipo'])) == 1 else 0,
                                                    tipoparticipante_id=2 if int(int(f.cleaned_data['tipo'])) == 1 else 5
                                                    )
                    docente.save(request)

                    # Guardo los niveles
                    for niv in f.cleaned_data["nivel"]:
                        docente.nivel.add(niv)

                    request.session['pestaña'] = 'docentesParticipantes'
                    request.session['vista'] = '2'

                    log(u'%s adicionó participante docente: %s' % (persona, docente), request, "add")
                    return JsonResponse({"result": "ok", "titulo": "Proceso exitoso!!!", "mensaje": u"Registro guardado con éxito", "showSwal": True})
                else:
                    msg = ",".join([k + ': ' + v[0] for k, v in f.errors.items()])
                    return JsonResponse({"result": "bad", "titulo": "Error", "mensaje": u"Error al guardar los datos. [%s]" % msg, "showSwal": "True", "swalType": "error"})
            except Exception as ex:
                msg = ex.__str__()
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "titulo": "Error", "mensaje": u"Error al guardar los datos. [%s]" % msg, "showSwal": "True", "swalType": "error"})

        if action == 'editparticipantesdocentes':
            try:
                es_externo = False
                externo = None
                personae = None

                f = ParticipanteProfesorVinculacionForm(request.POST)
                if f.is_valid():
                    tipodocente = int(request.POST['tipodocente'])

                    # Verificar que no esté repetido el integrante
                    if tipodocente == 1:
                        if ParticipantesMatrices.objects.filter(proyecto_id=request.POST['idproyecto'], profesor_id=f.cleaned_data['profesor'], status=True).exclude(pk=request.POST['id']).exists():
                            return JsonResponse({"result": "bad", "titulo": "Atención!!!", "mensaje": "El docente ya se encuentra inscrito en el proyecto", "showSwal": "True", "swalType": "warning"})
                    else:
                        # Consulto la persona
                        personae = Persona.objects.get(pk=f.cleaned_data['profesor'])

                        # Verifico si tiene perfil externo
                        if Externo.objects.filter(persona=personae, status=True).exists():
                            es_externo = True
                            externo = Externo.objects.get(persona=personae, status=True)

                            if ParticipantesMatrices.objects.filter(proyecto_id=request.POST['idproyecto'], externo_id=externo.id, status=True).exclude(pk=request.POST['id']).exists():
                                return JsonResponse({"result": "bad", "titulo": "Atención!!!", "mensaje": "El docente ya se encuentra inscrito en el proyecto", "showSwal": "True", "swalType": "warning"})

                    docente = ParticipantesMatrices.objects.get(pk=request.POST['id'])

                    if tipodocente == 1:
                        docente.profesor = Profesor.objects.get(pk=f.cleaned_data['profesor'])
                        docente.horas = f.cleaned_data['horas']
                        docente.save(request)

                        docente.nivel.clear()
                        for niv in f.cleaned_data["nivel"]:
                            docente.nivel.add(niv)
                    else:
                        # Si no existe externo, se debe crear ese perfil, Docentes externos
                        if tipodocente == 2:
                            if es_externo is False:
                                # Guardo externo
                                externo = Externo(
                                    persona=personae,
                                    nombrecomercial='',
                                    institucionlabora='',
                                    cargodesempena=''
                                )
                                externo.save(request)

                        docente.externo = externo
                        docente.save(request)

                    request.session['pestaña'] = 'docentesParticipantes'
                    request.session['vista'] = '2'

                    log(u'%s editó participante docente: %s' % (persona, docente), request, "edit")
                    return JsonResponse({"result": "ok", "titulo": "Proceso exitoso!!!", "mensaje": u"Registro actualizado con éxito", "showSwal": True})
                else:
                    msg = ",".join([k + ': ' + v[0] for k, v in f.errors.items()])
                    return JsonResponse({"result": "bad", "titulo": "Error", "mensaje": u"Error al guardar los datos. [%s]" % msg, "showSwal": "True", "swalType": "error"})
            except Exception as ex:
                msg = ex.__str__()
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "titulo": "Error", "mensaje": u"Error al guardar los datos. [%s]" % msg, "showSwal": "True", "swalType": "error"})

        if action == 'adddocenteexterno':
            try:
                if not 'idp' in request.POST:
                    return JsonResponse({"result": "bad", "titulo": "Error", "mensaje": u"Error al enviar los datos", "showSwal": "True", "swalType": "error"})

                f = DocenteExternoForm(request.POST)

                if f.is_valid():
                    # Verifica si existe la persona
                    if Persona.objects.values('id').filter(Q(cedula=f.cleaned_data['cedula'])|Q(pasaporte=f.cleaned_data['cedula']), status=True).exists():
                        return JsonResponse({"result": "bad", "titulo": "Atención!!!", "mensaje": "La persona ya está registrada en la base de datos", "showSwal": "True", "swalType": "warning"})

                    if f.cleaned_data['pasaporte']:
                        if Persona.objects.values('id').filter(Q(cedula=f.cleaned_data['pasaporte'])|Q(pasaporte=f.cleaned_data['pasaporte']), status=True).exists():
                            return JsonResponse({"result": "bad", "titulo": "Atención!!!", "mensaje": "La persona ya está registrada en la base de datos", "showSwal": "True", "swalType": "warning"})

                    # proyecto = ProyectosInvestigacion.objects.get(pk=int(encrypt(request.POST['idp'])))
                    proyecto = ProyectosInvestigacion.objects.get(pk=request.POST['idp'])

                    # Guardo la persona
                    personaexterna = Persona(
                        nombres=f.cleaned_data['nombres'],
                        apellido1=f.cleaned_data['apellido1'],
                        apellido2=f.cleaned_data['apellido2'],
                        cedula=f.cleaned_data['cedula'],
                        pasaporte=f.cleaned_data['pasaporte'],
                        nacimiento=f.cleaned_data['nacimiento'],
                        sexo=f.cleaned_data['sexo'],
                        nacionalidad=f.cleaned_data['nacionalidad'],
                        email=f.cleaned_data['email'].strip().lower(),
                        telefono=f.cleaned_data['telefono']
                    )
                    personaexterna.save(request)

                    # Guardo externo
                    externo = Externo(
                        persona=personaexterna,
                        nombrecomercial='',
                        institucionlabora=f.cleaned_data['institucionlabora'],
                        cargodesempena=f.cleaned_data['cargodesempena']
                    )
                    externo.save(request)

                    #Creo perfil externo
                    personaexterna.crear_perfil(externo=externo)
                    personaexterna.mi_perfil()
                    log(u'Agregó persona externa: %s' % (personaexterna), request, "add")

                    # Asigno la persona externa al proyecto con la función que cumple
                    integranteproyecto = ParticipantesMatrices(
                        matrizevidencia_id=2,
                        proyecto_id=request.POST['idp'],
                        profesor_id=None,
                        externo_id=externo.id,
                        horas=0,
                        tipoparticipante_id=5
                    )
                    integranteproyecto.save(request)

                    log(u'%s agregó integrante al proyecto: %s - %s' % (persona, proyecto, personaexterna), request, "add")
                    return JsonResponse({"result": "ok", "titulo": "Proceso exitoso!!!", "mensaje": u"Registro guardado con éxito", "showSwal": True})
                else:
                    msg = ",".join([k + ': ' + v[0] for k, v in f.errors.items()])
                    return JsonResponse({"result": "bad", "titulo": "Error", "mensaje": u"Error al guardar los datos. [%s]" % msg, "showSwal": "True", "swalType": "error"})
            except Exception as ex:
                msg = ex.__str__()
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "titulo": "Error", "mensaje": u"Error al guardar los datos. [%s]" % msg, "showSwal": "True", "swalType": "error"})

        if action == 'addinforme':
            try:
                f = FechaInformeForm(request.POST, request.FILES)
                if f.is_valid():
                    if 'archivo' in request.FILES:
                        newfile = request.FILES['archivo']
                    proyecto = ProyectosInvestigacion.objects.get(pk=request.POST['id'], tipo=1, status=True)
                    if InformeMarcoLogicoProyectosInvestigacion.objects.filter(proyectovinculacion=proyecto,
                                                                               fecha=f.cleaned_data['fecha'],
                                                                               status=True).exists():
                        return JsonResponse({"result": "r", "mensaje": u"Ya existe un informe en esa fecha."})
                    informemarcologicoproyectosinvestigacion = InformeMarcoLogicoProyectosInvestigacion(
                        proyectovinculacion=proyecto,
                        fecha=f.cleaned_data['fecha'], archivo=newfile,descripcion=f.cleaned_data['descripcion'].upper())
                    informemarcologicoproyectosinvestigacion.save(request)

                    marcologicoproyectosinvestigacionfines = proyecto.marcologicoproyectosinvestigacion_set.filter(
                        status=True, tipo=0).order_by('id')
                    marcologicoproyectosinvestigacionpropositos = proyecto.marcologicoproyectosinvestigacion_set.filter(
                        status=True, tipo=1).order_by('id')
                    marcologicoproyectosinvestigacioncomponentes = proyecto.marcologicoproyectosinvestigacion_set.filter(
                        status=True, tipo=2).order_by('id')
                    marcologicoproyectosinvestigacionacciones = proyecto.marcologicoproyectosinvestigacion_set.filter(
                        status=True, tipo=3).order_by('id')
                    for marcologicoproyectosinvestigacionfin in marcologicoproyectosinvestigacionfines:
                        detalleinforme = DetalleInforme(
                            informemarcologicoproyectosinvestigacion=informemarcologicoproyectosinvestigacion,
                            marcologicoproyectosinvestigacion=marcologicoproyectosinvestigacionfin)
                        detalleinforme.save(request)

                    for marcologicoproyectosinvestigacionproposito in marcologicoproyectosinvestigacionpropositos:
                        detalleinforme = DetalleInforme(
                            informemarcologicoproyectosinvestigacion=informemarcologicoproyectosinvestigacion,
                            marcologicoproyectosinvestigacion=marcologicoproyectosinvestigacionproposito)
                        detalleinforme.save(request)

                    for marcologicoproyectosinvestigacioncomponente in marcologicoproyectosinvestigacioncomponentes:
                        detalleinforme = DetalleInforme(
                            informemarcologicoproyectosinvestigacion=informemarcologicoproyectosinvestigacion,
                            marcologicoproyectosinvestigacion=marcologicoproyectosinvestigacioncomponente)
                        detalleinforme.save(request)

                    for marcologicoproyectosinvestigacionaccione in marcologicoproyectosinvestigacionacciones:
                        detalleinforme = DetalleInforme(
                            informemarcologicoproyectosinvestigacion=informemarcologicoproyectosinvestigacion,
                            marcologicoproyectosinvestigacion=marcologicoproyectosinvestigacionaccione)
                        detalleinforme.save(request)
                    log(u'Adiciono informe: %s' % informemarcologicoproyectosinvestigacion, request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        if action == 'addparticipantesestudiantes':
            try:
                f = ParticipanteEstudianteForm(request.POST)
                if f.is_valid():
                    datainscripcion = Inscripcion.objects.get(pk=int(f.cleaned_data['inscripcion']))
                    if ParticipantesMatrices.objects.filter(proyecto_id=request.POST['id'], inscripcion__persona=datainscripcion.persona, status=True).exists():
                        return JsonResponse({"result": "r", "mensaje": u"Error el estudiante ya se encuentra inscrito en el proyecto."})
                    programas = ParticipantesMatrices(matrizevidencia_id=2,
                                                      proyecto_id=request.POST['id'],
                                                      inscripcion_id=f.cleaned_data['inscripcion'],
                                                      horas=f.cleaned_data['horas']
                                                      )
                    programas.save(request)
                    log(u'Adiciono participante estudiante proyecto de investigación: %s' % programas, request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                     raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        if action == 'delparticipantesproyectos':
            try:
                nompersona = ''
                proyecto = ProyectosInvestigacion.objects.get(pk=request.POST['idproyecto'],tipo=1,status=True)
                participantes = ParticipantesMatrices.objects.filter(proyecto=proyecto,status=True)
                participantes1 = ProyectoVinculacionInscripcion.objects.filter(status=True,proyectovinculacion=proyecto)
                for participante1 in participantes1:
                    participante1.status=False
                    participante1.save(request)
                for participante in participantes:
                    participante.status = False
                    if participante.inscripcion:
                        nompersona = ' ESTUDIANTE ' + participante.inscripcion.persona.apellido1 + ' ' + participante.inscripcion.persona.apellido2 + ' ' + participante.inscripcion.persona.nombres
                    if participante.profesor:
                        nompersona = ' DOCENTE ' + participante.profesor.persona.apellido1 + ' ' + participante.profesor.persona.apellido2 + ' ' + participante.profesor.persona.nombres
                    if participante.administrativo:
                        nompersona = ' ADMINISTRATIVO ' + participante.administrativo.persona.apellido1 + ' ' + participante.administrativo.persona.apellido2 + ' ' + participante.administrativo.persona.nombres
                    participante.save(request)
                    log(u'Eliminación masiva de participantes de proyectos: %s - %s' % (proyecto, nompersona), request, "del")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al solicitar cupo."})

        if action == 'addparticipantesadministrativos':
            try:
                f = ParticipanteAdministrativoForm(request.POST)
                if f.is_valid():
                    if ParticipantesMatrices.objects.filter(proyecto_id=request.POST['id'],administrativo_id=f.cleaned_data['administrativo'],status=True).exists():
                        return JsonResponse({"result": "r","mensaje": u"Error personal administrativo ya se encuentra inscrito en el proyecto."})
                    programas = ParticipantesMatrices(matrizevidencia_id=2,
                                                      proyecto_id=request.POST['id'],
                                                      administrativo_id=f.cleaned_data['administrativo'],
                                                      horas=f.cleaned_data['horas']
                                                      )
                    programas.save(request)
                    log(u'Adiciono participante Administrativo programa de investigación: %s' % programas, request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                     raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'updatehoras':
            try:

                participantes = ParticipantesMatrices.objects.get(pk=request.POST['indi'])
                valor = int(request.POST['valor'])
                cupoanterior = participantes.horas
                participantes.horas = valor
                participantes.save(request)
                return JsonResponse({'result': 'ok', 'valor': participantes.horas})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad'})

        elif action == 'updatetipoparticipante':
            try:
                participantes = ParticipantesMatrices.objects.get(pk=request.POST['idparticipante'])
                tipopar = request.POST['tipoparticipante']
                tipoparanterior = participantes.tipoparticipante_id
                participantes.tipoparticipante_id = tipopar
                participantes.save(request)
                log(u'Editó tipo de participante de investigación: %s' % participantes, request, "add")
                return JsonResponse({'result': 'ok', 'valor': participantes.tipoparticipante_id})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad'})

        if action == 'addcarrerasproyectos':
            try:
                idproyecto = request.POST['idproyecto']
                carrerasproyectos = CarrerasProyecto.objects.filter(proyecto_id=int(idproyecto))
                carrerasproyectos.delete()
                lista = request.POST['listacarrerasproyecto']
                listacantidad = request.POST['listacantidadproyecto']
                if lista:
                    elementos = lista.split(',')
                    elementoscantidad = listacantidad.split(',')
                    i = 0
                    for elemento in elementos:
                        addcarrera = CarrerasProyecto(carrera_id=int(elemento),proyecto_id=int(idproyecto), cantidad=int(elementoscantidad[i]))
                        addcarrera.save(request)
                        log(u'Adiciono carreras de proyecto de investigación: %s' % addcarrera, request, "add")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        if action == 'addpresupuestoproyectos':
            try:
                idproyecto = request.POST['idproyecto']
                presupuesto = PresupuestosProyecto(anioejecucion=int(request.POST['anioejecucion']),
                                                   planificado=float(request.POST['planificado']),
                                                   ejecutado=float(request.POST['ejecutado']),
                                                   proyecto_id=int(idproyecto))
                presupuesto.save(request)
                log(u'Adiciono presupuesto de  proyectos de investigación: %s' % presupuesto, request, "add")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        if action == 'programapdf':
            try:
                data['programasvinculacion'] = programasvinculacion = ProgramasInvestigacion.objects.get(pk=request.POST['id'])
                programasvinculacioncampos = programasvinculacion.programasvinculacioncampos_set.filter(status=True)
                data['programasvinculacioncampos'] = None
                bandera=0
                if programasvinculacioncampos.exists():
                    bandera=1
                    data['programasvinculacioncampos'] = programasvinculacioncampos[0]
                data['bandera']=bandera
                return conviert_html_to_pdf(
                    'inv_vinculacion/programa.html',
                    {
                        'pagesize': 'A4',
                        'data': data,
                    }
                )
            except Exception as ex:
                pass

        if action == 'proyectopdf':
            try:
                data['proyectosinvestigacion'] = proyectosinvestigacion = ProyectosInvestigacion.objects.get(pk=request.POST['id'])
                participantesmatrices = proyectosinvestigacion.participantesmatrices_set.filter(status=True, tipoparticipante__id=1)
                data['participantesmatrices'] = ''
                data['coordinacion'] = ''
                data['decano'] = ''
                if participantesmatrices:
                    data['participantesmatrices'] = participantesmatrices[0].profesor.persona.nombre_titulo()
                    data['coordinacion'] = participantesmatrices[0].profesor.coordinacion
                    data['decano'] = participantesmatrices[0].profesor.coordinacion.responsable().persona.nombre_titulo()
                proyectosinvestigacioncampos=proyectosinvestigacion.proyectovinculacioncampos_set.filter(status=True)
                data['proyectosinvestigacioncampos'] = None
                if proyectosinvestigacioncampos.exists():
                    data['proyectosinvestigacioncampos'] = proyectosinvestigacioncampos[0]
                return conviert_html_to_pdf(
                    'inv_vinculacion/proyecto.html',
                    {
                        'pagesize': 'A4',
                        'data': data,
                    }
                )
            except Exception as ex:
                pass

        elif action == 'addevidenciaformato':
            try:
                f = EvidenciaFormatoForm(request.POST)
                if f.is_valid():
                    if Evidencia.objects.filter(nombre=f.cleaned_data['nombre'], matrizevidencia_id=2).exists():
                        return JsonResponse({"result": "bad", "mensaje": u"Registro Repetido."})
                    evidencia = Evidencia(nombre=f.cleaned_data['nombre'],
                                          matrizevidencia_id=2)
                    evidencia.save(request)
                    log(u'Adicionado formato evidencia: %s' % evidencia, request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                     raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        if action == 'editevidenciaformato':
            try:
                f = EvidenciaFormatoForm(request.POST)
                if f.is_valid():
                    evidencia = Evidencia.objects.get(pk=request.POST['id'])
                    evidencia.nombre = f.cleaned_data['nombre']
                    evidencia.save(request)
                    log(u'Edito formato evidencia: %s' % evidencia, request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                     raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        if action == 'addlineaprograma':
            try:
                programa = ProgramasInvestigacion.objects.get(pk=request.GET['id'])
                f = ProgramasVinculacionLineaForm(request.POST)
                if f.is_valid():
                    if LineaProgramasInvestigacion.objects.filter(programasinvestigacion=programa, lineainvestigacion=f.cleaned_data['lineainvestigacion'], sublineainvestigacion=f.cleaned_data['sublineainvestigacion'], status=True).exists():
                        return JsonResponse({"result": "bad", "mensaje": u"Registro Repetido."})
                    lineaprogramasinvestigacion = LineaProgramasInvestigacion(programasinvestigacion=programa,
                                                                              lineainvestigacion=f.cleaned_data['lineainvestigacion'],
                                                                              sublineainvestigacion=f.cleaned_data['sublineainvestigacion'])
                    lineaprogramasinvestigacion.save(request)
                    log(u'Adicionado linea de investigacion al programa: %s' % lineaprogramasinvestigacion, request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                     raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        if action == 'addareaprograma':
            try:
                programa = ProgramasInvestigacion.objects.get(pk=request.GET['id'])
                f = ProgramasVinculacionAreaForm(request.POST)
                if f.is_valid():
                    if AreaProgramasInvestigacion.objects.filter(programasinvestigacion=programa, areaconocimiento=f.cleaned_data['areaconocimiento'], subareaconocimiento=f.cleaned_data['subareaconocimiento'], subareaespecificaconocimiento=f.cleaned_data['subareaespecificaconocimiento'], status=True).exists():
                        return JsonResponse({"result": "bad", "mensaje": u"Registro Repetido."})
                    programasinvestigacionarea = AreaProgramasInvestigacion(programasinvestigacion=programa,
                                                                            areaconocimiento=f.cleaned_data['areaconocimiento'],
                                                                            subareaconocimiento=f.cleaned_data['subareaconocimiento'],
                                                                            subareaespecificaconocimiento=f.cleaned_data['subareaespecificaconocimiento'])
                    programasinvestigacionarea.save(request)
                    log(u'Adicionado area de conocimiento al programa: %s' % programasinvestigacionarea, request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                     raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        if action == 'dellineaprograma':
            try:
                lineaprogramasinvestigacion = LineaProgramasInvestigacion.objects.get(pk=request.POST['id'])
                log(u"Elimino linea de investigación en programa: %s" % lineaprogramasinvestigacion, request, "del")
                lineaprogramasinvestigacion.delete()
                return JsonResponse({"result": "ok", "id": periodo.id})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        if action == 'delareaprograma':
            try:
                areaprogramasinvestigacion = AreaProgramasInvestigacion.objects.get(pk=request.POST['id'])
                log(u"Elimino area de conocimeito en programa: %s" % areaprogramasinvestigacion, request, "del")
                areaprogramasinvestigacion.delete()
                return JsonResponse({"result": "ok", "id": periodo.id})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        if action == 'aprobacion_informe':
            try:

                idtarea = request.POST['id']
                detalle = DetalleCumplimiento.objects.get(pk=idtarea)
                if request.POST['est'] == '1':
                    detalle.aprobacion = True
                    detalle.aprobacion_adm = True
                    estado = 'aprobado'
                else:
                    detalle.aprobacion_adm = False
                    detalle.aprobacion = False
                    estado = 'rechazado'

                detalle.detalle_aprobacion = request.POST['obs']
                detalle.save(request)
                crono = Cronograma.objects.get(pk=detalle.tarea.pk)
                if detalle.tarea.avance() == float(detalle.tarea.cumplimiento):
                    crono.estado_finalizado = True
                else:
                    crono.estado_finalizado = False
                crono.save(request)

                if detalle.profesor:
                    notificacion = Notificacion(titulo=f"Evidencia de avance de proyecto {estado}",
                                                cuerpo=f"{persona.nombre_completo_inverso()} ha {estado} la evicencia de avance del proyecto de Servicio Comunitario.",
                                                destinatario=detalle.profesor.persona,
                                                url="/proyectovinculaciondocente?action=ejecucion&id=" + str(detalle.proyecto.pk),
                                                fecha_hora_visible=datetime.now() + timedelta(days=2),
                                                content_type=None,
                                                object_id=None,
                                                prioridad=1,
                                                app_label='sga')
                    notificacion.save()

                log(u'Aprobó avance de ejecución de proyectos de investigación docente: %s' % detalle, request, "edit")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"error": True, "mensaje": u"Error al guardar los datos."})

        elif action == 'aprobarconfiguracion':
            try:
                f = AprobacionInformeForm(request.POST)
                if f.is_valid():
                    informe = ConfiguracionInformeVinculacion.objects.get(pk=request.POST['id'])
                    if not ParticipantesMatrices.objects.filter(status=True,proyecto=informe.proyecto,tipoparticipante__pk=1, activo=True).exists():
                        raise NameError(u" Líder inactivo o no está asignado.")
                    es_informe_lider = informe.usuario_creacion.persona_set.first() == informe.proyecto.lider()
                    aprobacion = int(f.cleaned_data['aprobacion'])

                    if  t := informe.proyecto.get_tecnicoasociado():
                        if persona == t.persona and not es_informe_lider and not t.reemplaza_lider:
                            aprobacion = 5 if aprobacion == 3 else aprobacion

                    informe.aprobacion = aprobacion
                    informe.detalle_aprobacion = f.cleaned_data['detalle_aprobacion']
                    informe.personaaprueba = persona
                    es_informe_lider and migrar_evidencia_proyecto_vinculacion(request, informe, int(informe.aprobacion) - 1)

                    informe.save(request)
                    log(u'Aprobó la configuración de informe: %s' % informe, request, "edit")
                    return JsonResponse({"result": "ok", 'estado': int(informe.aprobacion)})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "error": True, "mensaje": u"Error al guardar los datos. %s" % ex.__str__()})

        if action == 'firmainformevinculacion':
            try:
                datas, datau, mensajeerror = None, None, ''
                c = ConfiguracionInformeVinculacion.objects.filter(id=request.POST.get('id')).first()
                file = c.archivo
                certificado = request.FILES["firma"]
                password = request.POST['palabraclave']
                razon = request.POST['razon'] if 'razon' in request.POST else ''

                words = "%s" % persona.nombre_completo()
                x, y, page = obtener_posicion_x_y_saltolinea(file.url, words)

                if not x or not y:
                    raise NameError(f"No se encontró la palabra {words} en el archivo. Por favor verifique que este nombre se encuentre en el apartado de firmas.")

                x, y = x + 30, y - 230

                try:
                    bytes_certificado, extension_certificado = certificado.read(), os.path.splitext(certificado.name)[1][1:]
                    datau = JavaFirmaEc(archivo_a_firmar=file, archivo_certificado=bytes_certificado, extension_certificado=extension_certificado, password_certificado=password, page=page, reason="Validar informe de vinculación", lx=x, ly=y).sign_and_get_content_bytes()
                except Exception as ex:
                    ...

                if not datau:
                    return JsonResponse({'result': 'bad', 'mensaje': "Problemas con la firma electrónica. %s" % datas if datas else ''})

                documento_a_firmar = io.BytesIO()
                documento_a_firmar.write(datau)
                documento_a_firmar.seek(0)

                nombrefile_ = file.name.split('/')[-1].split('.')[0]
                _name = generar_nombre(f'{c.usuario_creacion.username}_', 'file.pdf')

                c.archivo = DjangoFile(documento_a_firmar, _name)
                c.save(request)

                if c.aprobacion == 3:
                    valido, _, dict = verificarFirmasPDF(c.archivo)
                    if not dict or not dict.get('firmasValidas'):
                        raise NameError('No se firmó el archivo')

                    val, mensajeerror = migrar_evidencia_proyecto_vinculacion(request, c, 4)
                    log(u'%s firmo documento: %s' % (persona, nombrefile_), request, "add")
                return JsonResponse({'result': 'ok', 'id': c.pk, 'proyecto': c.proyecto.pk, 'para_revision': False, 'mensajeerror': mensajeerror})
            except Exception as ex:
                return JsonResponse({'result': 'bad', 'mensaje': f"Problemas de conexión con la plataforma Firma Ec, por favor intentalo más tarde. {ex}"})

        elif action == 'firmainformevinculacionmasivo':
            try:
                solicitudes = ConfiguracionInformeVinculacion.objects.filter(pk__in=request.POST.get('pks', '0').split(','), status=True)
                try:
                    datas, datau = None, None
                    certificado = request.FILES["firma"]
                    passfirma = request.POST['palabraclave']
                    razon = request.POST.get('razon', '')
                    extension_certificado = os.path.splitext(certificado.name)[1][1:]
                    bytes_certificado = certificado.read()
                    palabras = u"%s" % persona.nombre_completo()
                    msn = msnaux1 = msnaux2 = ''
                    contador_err = 0
                    contador_success = 0
                    for soli in solicitudes:
                        # es_lider, es_integ = soli.proyecto.lider() == soli.profesor, soli.profesor.persona == soli.profesor
                        documento_a_firmar = soli.archivo
                        name_documento_a_firmar, extension_documento_a_firmar = os.path.splitext(documento_a_firmar.name)
                        posx, posy, numpaginafirma = obtener_posicion_x_y_saltolinea(documento_a_firmar.url, palabras)
                        if posy:
                            posx, posy = posx + 30, posy - 230
                            try:
                                datau = JavaFirmaEc(archivo_a_firmar=documento_a_firmar,
                                                    archivo_certificado=bytes_certificado,
                                                    extension_certificado=extension_certificado,
                                                    password_certificado=passfirma,
                                                    page=numpaginafirma, reason=u"Validar informe de vinculación", lx=posx,
                                                    ly=posy).sign_and_get_content_bytes()
                            except Exception as x:
                                if str(x) == 'Certificado no es válido':
                                    raise NameError('Por favor asegúrese que la contraseña sea correcta y/o vuelva a intentarlo más tarde.')
                                contador_err += 1
                                msnaux2 = u' Problemas con la firma electrónica [%s]'%x
                                pass
                            if datau:
                                documento_a_firmar = io.BytesIO()
                                documento_a_firmar.write(datau)
                                documento_a_firmar.seek(0)
                                nombrefile_ = name_documento_a_firmar.__str__().split('/')[-1].replace('.pdf', '')
                                _name = f'{nombrefile_}_signed_{request.user.username}_' + '.pdf'
                                soli.archivo.save(_name, ContentFile(documento_a_firmar.read()))
                                if soli.aprobacion == 3: _, msj = migrar_evidencia_proyecto_vinculacion(request, soli, 4)
                                contador_success += 1
                                log(u'Firmo Documento: {}'.format(name_documento_a_firmar), request, "add")
                        else:
                            contador_err += 1
                            msnaux1 = u' Sin nombre (%s) en el apartado de firmas. '%(palabras)

                    if contador_success > 0:
                        msn = 'Documentos firmados con éxito.'
                    if contador_err > 0:
                        msn += ' Existieron inconvenientes con %s documento(s) que NO fueron firmados.%s%s' % (str(contador_err), msnaux1, msnaux2)

                    return JsonResponse({'result': 'ok', 'proyecto': solicitudes[0].proyecto.pk, 'mensajeerror': msn})

                except Exception as ex:
                    return JsonResponse({'result': 'bad', 'mensaje': "%s" % ex.__str__()})
            except Exception as ex:
                pass

        elif action == 'addcambioconfiguracion':
            try:
                f = ConfiguracionCambioForm(request.POST)
                if f.is_valid():
                    configuracion = ConfiguracionCambio(
                        proyecto=f.cleaned_data['proyecto'],
                        fecha_inicio=f.cleaned_data['fecha_inicio'],
                        fecha_fin=f.cleaned_data['fecha_fin'],
                        tipo=f.cleaned_data['tipo'])
                    configuracion.save(request)
                    if 'archivo' in request.FILES:
                        newfile = request.FILES['archivo']
                        newfile._name = generar_nombre("proyecto_", newfile._name)
                        configuracion.archivo = newfile
                        configuracion.save(request)
                    log(u'Adiciono un nuevo cambio de configuración: %s' % configuracion, request, "add")
                    return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"error": True, "mensaje": u"Error al guardar los datos."})

        elif action == 'editcambioconfiguracion':
            try:
                f = ConfiguracionCambioForm(request.POST)
                if f.is_valid():
                    idcambio = request.POST['id']
                    cambio = ConfiguracionCambio.objects.get(pk=idcambio)
                    cambio.proyecto = f.cleaned_data['proyecto']
                    cambio.fecha_inicio = f.cleaned_data['fecha_inicio']
                    cambio.fecha_fin = f.cleaned_data['fecha_fin']
                    cambio.tipo = f.cleaned_data['tipo']
                    cambio.save(request)
                    if 'archivo' in request.FILES:
                        newfile = request.FILES['archivo']
                        newfile._name = generar_nombre("proyecto_", newfile._name)
                        cambio.archivo = newfile
                        cambio.save(request)
                    log(u'Edito un cambio de configuración: %s' % cambio, request, "edit")
                    return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"error": True, "mensaje": u"Error al guardar los datos."})

        elif action == 'editfechagestion':
            try:
                f = FechaGestionForm(request.POST)
                if f.is_valid():
                    idcambio = request.POST['id']
                    cambio = ParticipantesMatrices.objects.get(pk=idcambio)
                    cambio.fecha_inicio = f.cleaned_data['fecha_inicio']
                    cambio.fecha_fin = f.cleaned_data['fecha_fin']
                    cambio.activo = f.cleaned_data['activo']
                    cambio.save(request)
                    log(u'Edito fecha de gestion a docente participante de proyecto vinculacion: %s' % cambio, request, "edit")
                    return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"error": True, "mensaje": u"Error al guardar los datos."})

        elif action == 'delcambioconfiguracion':
            try:
                cambio = ConfiguracionCambio.objects.get(id=request.POST['id'])
                cambio.status = False
                cambio.save(request)
                log(u'Elimino cambio de configuraión: %s' % cambio, request, "del")
                return JsonResponse({"error": False})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        elif action == 'cambioDocente':
            try:
                f = CambioDocenteVinculacionForm(request.POST)
                if f.is_valid():
                    proyecto = ProyectosInvestigacion.objects.get(pk=request.POST['id'])
                    responsable = f.cleaned_data['docente']
                    nuevo = f.cleaned_data['cambio']
                    docenteactual= ParticipantesMatrices.objects.get(status=True, proyecto = proyecto,profesor=responsable)
                    docenteactual.activo=False
                    docenteactual.save(request)

                    nuevo = ParticipantesMatrices(matrizevidencia_id=2,
                                                      proyecto=proyecto,
                                                      profesor=nuevo,
                                                      horas=0,
                                                      tipoparticipante=ParticipantesTipo.objects.get(pk=2)
                                                      )
                    nuevo.save(request)

                    crono = Cronograma.objects.filter(status=True, estado_finalizado=False,responsable= f.cleaned_data['docente'], proyecto = proyecto)
                    for cron in crono:
                        avance = DetalleCumplimiento.objects.filter(status=True, tarea=cron).exists()
                        if not avance:
                            cron.responsable.remove(responsable)
                            cron.responsable.add(f.cleaned_data['cambio'])
                            cron.save(request)
                    log(u'Realizo cambio de docente en proyecto vinculacion: %s' % proyecto, request, "add")
                    return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"error": True, "mensaje": u"Error al guardar los datos."})

        elif action == 'subirInforme':
            try:
                informe = ConfiguracionInformeVinculacion.objects.get(pk=request.POST['id'])
                if 'archivo' in request.FILES:
                    newfile = request.FILES['archivo']
                    newfile._name = generar_nombre("informe_", newfile._name)
                    informe.archivo = newfile
                informe.save(request)
                msj = 'Archivo actualizado'
                if informe.aprobacion == 3:
                    val, msj = migrar_evidencia_proyecto_vinculacion(request, informe, 4)
                notificacion = Notificacion(titulo="Informe mensual de avance validado",
                                            cuerpo=f"{persona.nombre_completo_minus()} ha cargado su informe mensual de avance validado del proyecto {informe.proyecto.__str__().title()}",
                                            destinatario=informe.profesor.persona,
                                            url="/proyectovinculaciondocente?action=configurarinforme&id=" + str(informe.proyecto.pk),
                                            fecha_hora_visible=datetime.now() + timedelta(days=2),
                                            content_type=None,
                                            object_id=None,
                                            prioridad=1,
                                            app_label='sga')
                notificacion.save()
                log(u'Subió informe de vinculación: %s' % informe, request, "edit")
                return JsonResponse({"result": False, 'mensaje': msj}, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        elif action == 'addperiodoinscripcion':
            try:
                f = PeriodoInscripcionFrom(request.POST)
                if f.is_valid():
                    periodo = PeriodoInscripcionVinculacion(
                        periodo=f.cleaned_data['periodo'],
                        proyecto = ProyectosInvestigacion.objects.get(pk=request.POST['proyecto']),
                        observacion= f.cleaned_data['observacion'],
                        fechainicio= f.cleaned_data['fechainicio'],
                        fechafin= f.cleaned_data['fechafin'],
                    )
                    periodo.save(request)

                    carreras = CarrerasParticipantes.objects.filter(status=True, proyecto__pk= request.POST['proyecto'])
                    for carr in carreras:
                        carrera = CarreraInscripcionVinculacion(
                            periodo = periodo,
                            carrera = carr,
                        )
                        carrera.save(request)

                    log(u'Añadió periodo de inscripcion de proyecto de vinculacion: %s' % periodo, request, "add")
                    return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"error": True, "mensaje": u"Error al guardar los datos."})

        elif action == 'editperiodoinscripcion':
            try:
                f = PeriodoInscripcionFrom(request.POST)
                if f.is_valid():
                    periodo = PeriodoInscripcionVinculacion.objects.get(pk=request.POST['id'])

                    periodo.periodo = f.cleaned_data['periodo']
                    periodo.observacion = f.cleaned_data['observacion']
                    periodo.fechainicio = f.cleaned_data['fechainicio']
                    periodo.fechafin = f.cleaned_data['fechafin']
                    periodo.save(request)

                    log(u'Editó periodo de inscripcion de proyecto de vinculacion: %s' % periodo, request, "edit")
                    return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"error": True, "mensaje": u"Error al guardar los datos."})

        elif action == 'actualizacupo':
            try:
                carrera = CarreraInscripcionVinculacion.objects.get(pk=request.POST['carrera'])
                carrera.cupos = int(request.POST['cupo'])
                carrera.save(request)
                return JsonResponse({'result': 'ok', 'cupo': carrera.cupos})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad'})

        elif action == 'aprobarperiodo':
            try:
                idperiodo = request.POST['id']
                detalle = PeriodoInscripcionVinculacion.objects.get(pk=idperiodo)
                detalle.aprobado = True
                detalle.save(request)

                log(u'Aprobó periodo de inscripción de estudiante: %s' % detalle, request, "edit")
                return JsonResponse({"error": False})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"error": True, "mensaje": u"Error al guardar los datos."})

        elif action == 'delperiodo':
            try:
                idperiodo = request.POST['id']
                detalle = PeriodoInscripcionVinculacion.objects.get(pk=idperiodo)
                detalle.status = False
                detalle.save(request)

                log(u'Eliminó periodo de inscripción de estudiantes: %s' % detalle, request, "del")
                return JsonResponse({"error": False})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"error": True, "mensaje": u"Error al guardar los datos."})

        elif action == 'aprobarsolicitudproyecto':
            try:
                participante = ProyectoVinculacionInscripcion.objects.get(pk=request.POST['id'])
                participante.estado = 2
                participante.save(request)

                if not ParticipantesMatrices.objects.filter(proyecto=participante.proyectovinculacion,inscripcion=participante.inscripcion, status=True).exists():
                    programas = ParticipantesMatrices(matrizevidencia_id=2,
                                                      proyecto=participante.proyectovinculacion,
                                                      inscripcion=participante.inscripcion,
                                                      horas=0,
                                                      preinscripcion=participante,
                                                      )
                    programas.save(request)

                    saludo = 'Estimada ' if participante.inscripcion.persona.sexo_id == 1 else 'Estimado '
                    notificacion = Notificacion(
                        titulo=f"Estado de solicitud de participación en proyectos de vinculación",
                        cuerpo=f"{saludo}  {participante.inscripcion.persona.nombre_completo_inverso()}, su preinscripción al proyecto de vinculación {programas.proyecto.nombre} ha sido aprobada",
                        destinatario=participante.inscripcion.persona,
                        url="/alu_proyectovinculacion?panel=3",
                        fecha_hora_visible=datetime.now() + timedelta(days=2),
                        content_type=None,
                        object_id=None,
                        prioridad=1,
                        app_label='sga')
                    notificacion.save()

                    asunto = "Solicitud de Proyecto de Vinculación aprobada."
                    mensaje = "ha sido revisado su informe"
                    template = "emails/aceptacion_preinscripcion_vinc.html"
                    datos = {'sistema': u'SGA - UNEMI',
                             'mensaje': mensaje,
                             'fecha': datetime.now().date(),
                             'hora': datetime.now().time(),
                             'saludo': 'Estimada' if programas.inscripcion.persona.sexo_id == 1 else 'Estimado',
                             'estudiante': programas.inscripcion.persona.nombre_completo_inverso(),
                             'proyecto': programas.proyecto.nombre,
                             'autoridad2': '',
                             't': miinstitucion()
                             }
                    email = programas.inscripcion.persona.lista_emails_envio()

                    send_html_mail(asunto, template, datos, email, [], cuenta=variable_valor('CUENTAS_CORREOS')[0])
                    # log(u'Aprobo solicitud proyecto vinparticipantesproyectosculacion docente: %s' % participante, request, "edit")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        elif action == 'editcarrera':
            try:
                # inscrip_actividad = InscripcionActividadConvalidacionPPV.objects.get(pk=request.POST['id'])

                id = request.POST['id_parti']

                inscrip_participantes = ParticipantesMatrices.objects.get(pk=id)

                f = CambioInscripcionVinculacionForm(request.POST)
                if f.is_valid():

                    inscripciondestino = Inscripcion.objects.get(persona=inscrip_participantes.inscripcion.persona,
                                                                 status=True,
                                                                 perfilusuario__status=True,
                                                                 perfilusuario__visible=True,
                                                                 carrera=f.cleaned_data['carreradestino'])


                    inscrip_actividad = inscrip_participantes.inscripcion
                    inscrip_participantes.inscripcion = inscripciondestino
                    inscrip_actividad.save(request)

                    inscrip_participantes.save(request)

                    log(u'%s cambió de inscripción al proyecto [ %s ] de %s a %s' % (persona, inscrip_participantes.proyecto, inscrip_participantes.inscripcion.carrera, inscripciondestino.carrera), request,"edit")

                    return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al actualizar los datos."})

        elif action == 'deletesolicitudproyecto':
            try:
                participante = ProyectoVinculacionInscripcion.objects.get(pk=request.POST['id'])

                proyecto  =  ProyectosInvestigacion.objects.get(pk=participante.proyectovinculacion.id)
                proyecto.saldoo = proyecto.saldoo + 1
                proyecto.save(request)
                #log(u'Elimino solicitud proyecto vinculacion docente: %s' % participante, request, "del")
                participante.delete()
                #participante.saldo(request)
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        elif action == 'definirregistrohoras':
            try:
                f = FechasRegistroHorasForm(request.POST)
                if f.is_valid():
                    proyecto = ProyectosInvestigacion.objects.get(pk=request.POST['id'])
                    for matriz in ParticipantesMatrices.objects.filter(status=True, proyecto=proyecto):
                        matriz.registrohorasdesde = f.cleaned_data['fechadesde']
                        matriz.registrohorashasta = f.cleaned_data['fechahasta']
                        matriz.save(request)
                    log(u'Guardó la definición de fechas de registro de horas: %s' % matriz, request, "edit")
                    return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'addhabilitacioninforme':
            try:
                with transaction.atomic():
                    proyecto = ProyectosInvestigacion.objects.get(pk=request.POST['proyectoid'])
                    participantes = ParticipantesMatrices.objects.filter(status=True, proyecto=proyecto, matrizevidencia_id=2, horas=0)
                    estudiantes = participantes.filter(inscripcion__in=participantes.values_list('inscripcion_id', flat=True))
                    form = InformeProyectoVinculacionForm(request.POST, request.FILES)
                    if form.is_valid():
                        informe = InformesProyectoVinculacionDocente(proyecto=proyecto,
                                                                     nombre=form.cleaned_data['nombre'].upper(),
                                                                     descripcion=form.cleaned_data['descripcion'],
                                                                     flimite=form.cleaned_data['flimite'])
                        # informe.save(request)
                        if 'formato' in request.FILES:
                            newfile = request.FILES['formato']
                            newfile._name = generar_nombre(informe.nombre_input(), newfile._name)
                            informe.formato = newfile

                        log(u'Adicionó Habilitación de Informe: %s' % informe, request, "add")

                        asunto = "Habilitación de informe de Proyecto de Servicio comunitario."
                        mensaje = "ha sido habilitada la carga del informe"
                        template = "emails/notificarhabilitacioninformevinculacion.html"
                        datos = {'sistema': u'SGA - UNEMI',
                                 'mensaje': mensaje,
                                 'fecha': datetime.now().date(),
                                 'hora': datetime.now().time(),
                                 'informe': informe,
                                 'proyecto': informe.proyecto.nombre,
                                 'autoridad2': '',
                                 't': miinstitucion()
                                 }

                        lista_correos_estudiantes = []
                        for estudiante in estudiantes:
                            lista_correos_estudiantes.append(estudiante.inscripcion.persona.emailinst)

                        send_html_mail(asunto, template, datos, lista_correos_estudiantes, [], cuenta=variable_valor('CUENTAS_CORREOS')[0])
                        informe.notificado = True
                        informe.save(request)
                        return JsonResponse({"result": False}, safe=False)
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({"result": True, "mensaje": "Complete los datos requeridos."}, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        elif action == 'editarhabilitacioninforme':
            try:
                with transaction.atomic():
                    informe = InformesProyectoVinculacionDocente.objects.get(pk=request.POST['id'])
                    f = InformeProyectoVinculacionForm(request.POST, request.FILES)
                    if f.is_valid():
                        informe.nombre = f.cleaned_data['nombre'].upper()
                        informe.descripcion = f.cleaned_data['descripcion']
                        informe.flimite = f.cleaned_data['flimite']
                        if 'formato' in request.FILES:
                            newfile = request.FILES['formato']
                            newfile._name = generar_nombre(informe.nombre_input(), newfile._name)
                            informe.formato = newfile
                        informe.save(request)
                        log(u'Modificó Actividad Extracurricular: %s' % informe, request, "edit")
                        return JsonResponse({"result": False}, safe=False)
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({"result": True, "form": [{k: v[0]} for k, v in f.errors.items()],
                                             "mensaje": "Complete los datos requeridos."}, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        elif action == 'eliminarhabilitacioninforme':
            try:
                informe = InformesProyectoVinculacionDocente.objects.get(pk=int(request.POST['id']))
                informe.status = False
                informe.save(request)
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'addhabilitacionhoras':
            try:
                with transaction.atomic():
                    proyecto = ProyectosInvestigacion.objects.get(pk=request.POST['proyectoid'])
                    convocatoria = PeriodoInscripcionVinculacion.objects.get(pk=int(request.POST['convocatoria']))
                    form = HabilitacionRegistroHorasVinculacionForm(request.POST, request.FILES)
                    if form.is_valid():
                        habilitacion = HabilitacionesHorasParticipantesVinculacion(proyecto=proyecto,
                                                                     convocatoria=convocatoria,
                                                                     registrohorasdesde=form.cleaned_data['fechadesde'],
                                                                     registrohorashasta=form.cleaned_data['fechahasta'])
                        if 'formato' in request.FILES:
                            newfile = request.FILES['formato']
                            newfile._name = generar_nombre("informe_", newfile._name)
                            habilitacion.formato = newfile

                        log(u'Adicionó Habilitación de Horas: %s' % habilitacion, request, "add")
                        habilitacion.save(request)
                        return JsonResponse({"result": False}, safe=False)
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({"result": True, "mensaje": "Complete los datos requeridos."}, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        elif action == 'editarhabilitacionhoras':
            try:
                with transaction.atomic():
                    # proyecto = ProyectosInvestigacion.objects.get(pk=request.POST['proyectoid'])
                    habilitacion = HabilitacionesHorasParticipantesVinculacion.objects.get(pk=request.POST['id'])
                    convocatoria = PeriodoInscripcionVinculacion.objects.get(pk=int(request.POST['convocatoria']))
                    f = HabilitacionRegistroHorasVinculacionForm(request.POST, request.FILES)
                    if f.is_valid():
                        habilitacion.convocatoria = convocatoria
                        habilitacion.registrohorasdesde = f.cleaned_data['fechadesde']
                        habilitacion.registrohorashasta = f.cleaned_data['fechahasta']
                        if 'formato' in request.FILES:
                            newfile = request.FILES['formato']
                            newfile._name = generar_nombre("informe_", newfile._name)
                            habilitacion.formato = newfile
                        habilitacion.save(request)
                        log(u'Modificó Actividad Extracurricular: %s' % habilitacion, request, "edit")
                        return JsonResponse({"result": False}, safe=False)
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({"result": True, "form": [{k: v[0]} for k, v in f.errors.items()],
                                             "mensaje": "Complete los datos requeridos."}, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        elif action == 'eliminarhabilitacionhoras':
            try:
                habilitacion = HabilitacionesHorasParticipantesVinculacion.objects.get(pk=int(request.POST['id']))
                habilitacion.status = False
                habilitacion.save(request)
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'addinscritomanual':
            try:
                idestudiante, id = request.POST['idestudiante'], request.POST['id']
                inscripcion = Inscripcion.objects.get(pk=idestudiante)
                periodoinscripcion = PeriodoInscripcionVinculacion.objects.get(pk=id)
                carrera = inscripcion.carrera
                proyectosinvestigacion = ProyectosInvestigacion.objects.get(pk=periodoinscripcion.proyecto.pk)
                cupos = CarreraInscripcionVinculacion.objects.get(carrera__carrera=carrera,periodo=periodoinscripcion).diferencia()
                if ItinerariosVinculacionMalla.objects.filter(status=True,malla__carrera=carrera).exists():
                    nivel_itinerario = ItinerariosVinculacionMalla.objects.filter(status=True,malla__carrera=carrera).values_list('nivel', flat=True).first()
                else:
                    nivel_itinerario = 4
                matricula=inscripcion.matricula_periodo(periodo)
                if matricula:
                    nivel = matricula.nivelmalla.orden
                else:
                    nivel = inscripcion.mi_nivel().nivel.orden
                if not inscripcion.perfil_inscripcion():
                    return JsonResponse({"result": False, "mensaje": u"La inscripción no se encuentra activa"})

                if ProyectoVinculacionInscripcion.objects.filter(status=True, periodo__periodo=periodo,inscripcion=inscripcion, estado__in=[1,2]).exists():
                    return JsonResponse({"result": False, "mensaje": u"Ya cuenta con una inscripción activa en este periodo"})
                if ProyectoVinculacionInscripcion.objects.filter(status=True, proyectovinculacion=proyectosinvestigacion,inscripcion=inscripcion).exists():
                    return JsonResponse({"result": False, "mensaje": u"Ya tiene una inscripción activa en este proyecto"})
                if ParticipantesMatrices.objects.filter(status=True, proyecto=proyectosinvestigacion,inscripcion=inscripcion).exists():
                    return JsonResponse({"result": False, "mensaje": u"Ya se encuentra registrado en este proyecto"})
                if ParticipantesMatrices.objects.filter(status=True, estado = 0, inscripcion = inscripcion, actividad__isnull = True).exists():
                    return JsonResponse({"result": False, "mensaje": u"Tiene aún una participación en proceso"})
                if cupos==0:
                    return JsonResponse({"result": False, "mensaje": u"Cupo no disponible"})

                proyectovinculacioninscripcion = ProyectoVinculacionInscripcion(
                    proyectovinculacion=proyectosinvestigacion,
                    inscripcion=inscripcion,
                    periodo=periodoinscripcion
                )
                proyectovinculacioninscripcion.save(request)

                if inscripcion.carrera.modalidad == 3:
                    matricula = inscripcion.matricula_periodo(periodo)
                    if materias_asignada_vinculacion := obtener_materia_asignada_vinculacion_por_nivel_v2(matricula.id):
                        if pdf := materias_asignada_vinculacion.actacompromisovinculacion:
                            extradetalle = ExtraProyectoVinculacionInscripcion(proyectoinscripcion=proyectovinculacioninscripcion, actacompromisovinculacion=pdf)
                            extradetalle.save(request)

                            # APROBADO AUTOMATICO
                            proyectovinculacioninscripcion.estado = 2
                            proyectovinculacioninscripcion.save(request)

                            if not ParticipantesMatrices.objects.filter(proyecto=proyectovinculacioninscripcion.proyectovinculacion, inscripcion=proyectovinculacioninscripcion.inscripcion, status=True).exists():
                                programas = ParticipantesMatrices(matrizevidencia_id=2,
                                                                  proyecto=proyectovinculacioninscripcion.proyectovinculacion,
                                                                  inscripcion=proyectovinculacioninscripcion.inscripcion,
                                                                  horas=0,
                                                                  preinscripcion=proyectovinculacioninscripcion
                                                                  )
                                programas.save(request)

                                saludo = 'Estimada ' if proyectovinculacioninscripcion.inscripcion.persona.sexo_id == 1 else 'Estimado '
                                notificacion = Notificacion(
                                    titulo=f"Estado de solicitud de participación en proyectos de vinculación",
                                    cuerpo=f"{saludo}  {proyectovinculacioninscripcion.inscripcion.persona.nombre_completo_inverso()}, su preinscripción al proyecto de vinculación {programas.proyecto.nombre} ha sido aprobada automaticamente por el sistema.",
                                    destinatario=proyectovinculacioninscripcion.inscripcion.persona,
                                    url="/alu_proyectovinculacion?panel=3",
                                    fecha_hora_visible=datetime.now() + timedelta(days=2),
                                    content_type=None,
                                    object_id=None,
                                    prioridad=1,
                                    app_label='sga')
                                notificacion.save()
                            # FIN APROBADO AUTOMATICO

                log(u'%s agregó inscripción en el proyecto de vinculación: %s' % (persona, proyectosinvestigacion), request, "add")

                return JsonResponse({"result": True})
            except Exception as ex:
                return JsonResponse({"result": False, 'mensaje': str(ex)})

        elif action == 'editcarrera':
            try:
                participante = ParticipantesMatrices.objects.get(pk=request.POST['id'])
                f = CambioCarreraVinculacionForm(request.POST)
                if f.is_valid():
                    inscripciondestino = Inscripcion.objects.get(persona=participante.inscripcion.persona,
                                                                 status=True,
                                                                 perfilusuario__status=True,
                                                                 perfilusuario__visible=True,
                                                                 carrera=f.cleaned_data['carreradestino'])

                    inscripcionant = f.cleaned_data['carreraactual']
                    participante.inscripcion = inscripciondestino
                    participante.save(request)

                    log(u'%s cambió carrera e itinerario a la práctica [ %s ] de %s a %s' % (persona, participante, inscripcionant, participante.inscripcion.carrera), request,"edit")

                    return JsonResponse({"result": "ok"})
                else:
                    errorformulario = f._errors
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'registrarhoras':
            try:
                participante = ParticipantesMatrices.objects.get(pk=int(encrypt(request.POST['id'])))

                participante.registrohorasdesde = datetime.strptime(request.POST['fechainicio'], '%d-%m-%Y')
                participante.registrohorashasta = datetime.strptime(request.POST['fechafinalizacion'], '%d-%m-%Y')
                participante.horas = request.POST['horas']
                participante.estado = request.POST['estado']
                participante.save(request)

                log(u'%s registró horas a %s ' % (persona, participante), request,"edit")

                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'edittecnicoasociado':
            try:
                tecnicoasociado = TecnicoAsociadoProyectoVinculacion.objects.get(id=request.POST.get('id'))
                f = TecnicoAsociadoProyectoVinculacionForm(request.POST)
                del f.fields['persona']
                if f.is_valid():
                    tecnicoasociado.cargo = f.cleaned_data.get('cargo')
                    tecnicoasociado.reemplaza_lider = f.cleaned_data.get('reemplaza_lider')
                    tecnicoasociado.save(request)
                    log(f"Editó configuración de técnico {tecnicoasociado.pk} asociado al proyecto {tecnicoasociado.proyecto.pk}", request, 'edit')
                    return JsonResponse({'result': True})
            except Exception as ex:
                pass

        elif action == 'reemplazaliderproyecto':
            try:
                ta = TecnicoAsociadoProyectoVinculacion.objects.get(id=request.POST.get('id'))
                ta.reemplaza_lider = not ta.reemplaza_lider
                ta.save(request)
                log(f"{persona} {'ACTIVO' if ta.reemplaza_lider else 'DESACTIVO'} la firma del lider del proyecto {ta.proyecto.pk}", request, 'edit')
                return JsonResponse({'result': True})
            except Exception as ex:
                pass

        elif action == 'addtecnicoasociado':
            try:
                f = TecnicoAsociadoProyectoVinculacionForm(request.POST)
                f.delpersona()

                hoy = datetime.now().date()

                if ultimo_activo := TecnicoAsociadoProyectoVinculacion.objects.filter(proyecto_id=request.POST.get('id'), activo=True, status=True).first():
                    ultimo_activo.activo = ultimo_activo.status = False
                    ultimo_activo.fechafin = hoy
                    ultimo_activo.personaelimina = persona
                    ultimo_activo.save(request)

                if not f.is_valid():
                    return JsonResponse({"result": "bad", "form": [{k: v[0]} for k, v in f.errors.items()]})

                tecnico = TecnicoAsociadoProyectoVinculacion(persona_id=request.POST.get('persona', None),
                                                             proyecto_id=request.POST.get('proyecto', None),
                                                             cargo=f.cleaned_data.get('cargo'),
                                                             fechainicio=hoy,
                                                             activo=True,
                                                             reemplaza_lider=f.cleaned_data.get('reemplaza_lider'))
                tecnico.save(request)
                log(f"{persona} adicionó técnico para revisión de informe al proyecto {tecnico.proyecto.pk}", request, 'add')
                return JsonResponse({'result': True})
            except Exception as ex:
                return JsonResponse({'result': False, 'mensaje': 'Error de conexión, %s' % ex.__str__()})

        elif action == 'subirarchivoinscripcion':
            try:
                with transaction.atomic():
                    directory = os.path.join(SITE_STORAGE, 'media', 'innovacion')
                    try:
                        os.stat(directory)
                    except:
                        os.mkdir(directory)
                    f = MasivoInscripcionVinculacionForm(request.POST)
                    if f.is_valid():
                        if not 'archivo' in request.FILES:
                            raise NameError('Debe subir un archivo válido')
                        nfile = request.FILES['archivo']
                        archivo = io.BytesIO(nfile.read())
                        nombre_archivo = 'Observaciones_' + str(nfile.name)
                        directory = os.path.join(MEDIA_ROOT, 'innovacion', nombre_archivo)
                        workbook = openpyxl.load_workbook(filename=archivo)
                        loes = workbook[workbook.sheetnames[0]]
                        totallista = loes.rows
                        contador = 0
                        hoja = workbook.active
                        ultima_columna  = hoja.max_column
                        borde = Border(left=Side(style='thin'),
                                       right=Side(style='thin'),
                                       top=Side(style='thin'),
                                       bottom=Side(style='thin'))
                        observacion = ''
                        ccedula = 2
                        ccorreo = 4
                        bandera = False
                        for rowx in range(1, loes.max_row + 1):
                            contador += 1
                            bandera = False
                            if contador == 1 :
                                ultima_celda_primera_fila = hoja.cell(row=1, column=ultima_columna)
                                siguiente_celda = hoja.cell(row=1, column=ultima_columna + 1)
                                siguiente_celda.font = copy(ultima_celda_primera_fila.font)
                                siguiente_celda.fill = copy(ultima_celda_primera_fila.fill)
                                siguiente_celda.alignment = alin(horizontal="center", vertical="center")
                                siguiente_celda.border = borde
                                hoja.column_dimensions[siguiente_celda.column_letter].width = 50
                                siguiente_celda.value = 'OBSERVACIONES'

                            else:
                                cedula = loes.cell(row=rowx, column=ccedula).value
                                insccurso = None
                                observacion = ''
                                if Inscripcion.objects.filter(persona__cedula=cedula, matricula__status = True, matricula__nivel__periodo_id = periodo.pk, status=True).exists():
                                    id = int(encrypt(request.POST['id']))
                                    periodoinscripcion = PeriodoInscripcionVinculacion.objects.get(pk=id)
                                    proyectosinvestigacion = ProyectosInvestigacion.objects.get(pk=periodoinscripcion.proyecto.pk)
                                    carreraid = CarrerasParticipantes.objects.values_list('carrera_id', flat=True).filter(status=True, proyecto=periodoinscripcion.proyecto)
                                    # inscripcion = Inscripcion.objects.filter(persona__cedula=cedula, matricula__status=True,matricula__nivel__periodo_id=periodo.pk,status=True, carrera__in = carreraid).first()
                                    inscripciones = Inscripcion.objects.filter(persona__cedula=cedula, matricula__status=True,matricula__nivel__periodo_id=periodo.pk,status=True, carrera__in = carreraid)
                                    if inscripciones:
                                        for ins in inscripciones:
                                            inscripcion = Inscripcion.objects.get(pk=ins.id)
                                            carrera = inscripcion.carrera

                                            if len(carreraid) > 0:
                                                if carrera.pk in carreraid:
                                                    if CarreraInscripcionVinculacion.objects.filter(carrera__carrera=carrera,periodo=periodoinscripcion).exists():
                                                        cupos = CarreraInscripcionVinculacion.objects.get(carrera__carrera=carrera,periodo=periodoinscripcion)
                                                        cupos_diferencia = cupos.diferencia()

                                                        if ItinerariosVinculacionMalla.objects.filter(status=True,malla__carrera=carrera).exists():
                                                            nivel_itinerario = ItinerariosVinculacionMalla.objects.filter(status=True, malla__carrera=carrera).values_list('nivel', flat=True).first()
                                                        else:
                                                            nivel_itinerario = 4
                                                        matricula = inscripcion.matricula_periodo(periodo)
                                                        if matricula:
                                                            nivel = matricula.nivelmalla.orden
                                                        else:
                                                            nivel = inscripcion.mi_nivel().nivel.orden
                                                        if not inscripcion.perfil_inscripcion():
                                                            observacion = "La inscripción no se encuentra activa"
                                                            bandera = True
                                                        else:
                                                            if ProyectoVinculacionInscripcion.objects.filter(status=True,periodo__periodo=periodo,inscripcion=inscripcion,estado__in=[1]).exists():
                                                                observacion = "Ya cuenta con una inscripción pendiente en este periodo - Codigo del proyecto: "+str(ProyectoVinculacionInscripcion.objects.filter(status=True,periodo__periodo=periodo,inscripcion=inscripcion,estado__in=[1]).first().proyectovinculacion.id)
                                                                bandera = True
                                                            elif ProyectoVinculacionInscripcion.objects.filter(status=True,periodo__periodo=periodo,inscripcion=inscripcion,estado__in=[2]).exists():
                                                                observacion = "Ya cuenta con una inscripción aprobada en este periodo - Codigo del proyecto: " +str(ProyectoVinculacionInscripcion.objects.filter(status=True,periodo__periodo=periodo,inscripcion=inscripcion,estado__in=[2]).first().proyectovinculacion.id)
                                                                bandera = True
                                                            elif ProyectoVinculacionInscripcion.objects.filter(status=True,proyectovinculacion=proyectosinvestigacion,inscripcion=inscripcion).exists():
                                                                observacion = "Ya tiene una inscripción activa en este proyecto"
                                                                bandera = True
                                                            elif ParticipantesMatrices.objects.filter(status=True,proyecto=proyectosinvestigacion,inscripcion=inscripcion).exists():
                                                                observacion = "Ya se encuentra registrado en este proyecto"
                                                                bandera = True
                                                            elif ParticipantesMatrices.objects.filter(status=True, estado=0,inscripcion=inscripcion,actividad__isnull=True).exists():
                                                                observacion = "Tiene aún una participación en proceso"
                                                                bandera = True

                                                        if carrera.pk in [128,132,133,130,134,135,127,129,126,131]:
                                                            if bandera == False:
                                                                if not ProyectoVinculacionInscripcion.objects.filter(status = True,proyectovinculacion=proyectosinvestigacion,inscripcion=inscripcion,periodo=periodoinscripcion).exists():
                                                                    if cupos_diferencia == 0:
                                                                        cupos.cupos += 1
                                                                        cupos.save()
                                                                    proyectovinculacioninscripcion = ProyectoVinculacionInscripcion(
                                                                        proyectovinculacion=proyectosinvestigacion,
                                                                        inscripcion=inscripcion,
                                                                        periodo=periodoinscripcion,
                                                                        usuario_modificacion = persona.usuario,
                                                                        fecha_creacion = datetime.now().date(),
                                                                        fecha_modificacion = datetime.now().date(),
                                                                        estado = 2
                                                                    )
                                                                    proyectovinculacioninscripcion.save(request)

                                                                    if inscripcion.carrera.modalidad == 3:
                                                                        matricula = inscripcion.matricula_periodo(periodo)
                                                                        if materias_asignada_vinculacion := obtener_materia_asignada_vinculacion_por_nivel_v2(matricula.id):
                                                                            if pdf := materias_asignada_vinculacion.actacompromisovinculacion:
                                                                                extradetalle = ExtraProyectoVinculacionInscripcion(proyectoinscripcion=proyectovinculacioninscripcion, actacompromisovinculacion=pdf)
                                                                                extradetalle.save(request)

                                                                ## -------------------------  INSCRIPCION ------------------------------
                                                                    personal_docente = ParticipantesMatrices.objects.filter(proyecto=proyectovinculacioninscripcion.proyectovinculacion, profesor_id__gt = 0, tipoparticipante_id = 1, status=True).first()
                                                                    if personal_docente:
                                                                        nombre_docente = personal_docente.profesor.persona.nombre_completo_inverso()
                                                                        correo_docente = personal_docente.profesor.persona.emailinst
                                                                    else:
                                                                        nombre_docente = 'DOCENTE NO ASIGNADO'
                                                                        correo_docente = '-'
                                                                    if not ParticipantesMatrices.objects.filter(proyecto=proyectovinculacioninscripcion.proyectovinculacion,inscripcion=proyectovinculacioninscripcion.inscripcion, status=True).exists():
                                                                        programas = ParticipantesMatrices(matrizevidencia_id=2,
                                                                                                          proyecto=proyectovinculacioninscripcion.proyectovinculacion,
                                                                                                          inscripcion=proyectovinculacioninscripcion.inscripcion,
                                                                                                          horas=0,
                                                                                                          preinscripcion=proyectovinculacioninscripcion,
                                                                                                          )
                                                                        programas.save()
                                                                        observacion = "Asignado"

                                                                        saludo = 'Estimada ' if proyectovinculacioninscripcion.inscripcion.persona.sexo_id == 1 else 'Estimado '
                                                                        notificacion = Notificacion(
                                                                            titulo=f"Asignación de Prácticas de Servicio Comunitario",
                                                                            cuerpo=f"{saludo}  {proyectovinculacioninscripcion.inscripcion.persona.nombre_completo_inverso()}, nos complace informarte que, has sido asignado para llevar a cabo tus prácticas de servicio comunitario en el proyecto: {programas.proyecto.nombre}, liderado por el docente {nombre_docente}. Por favor, comunícate con el docente responsable, {nombre_docente} al siguiente correo {correo_docente}, para coordinar los detalles específicos de los horarios y actividades a desarrollar dentro de este proceso.",
                                                                            destinatario=proyectovinculacioninscripcion.inscripcion.persona,
                                                                            url="/alu_proyectovinculacion?panel=3",
                                                                            fecha_hora_visible=datetime.now() + timedelta(days=2),
                                                                            content_type=None,
                                                                            object_id=None,
                                                                            prioridad=1,
                                                                            app_label='sga')
                                                                        notificacion.save()

                                                                        asunto = "Asignación de Prácticas de Servicio Comunitario"
                                                                        mensaje = "ha sido revisado su informe"
                                                                        template = "emails/aceptacion_preinscripcion_vinc_masiva.html"
                                                                        datos = {'sistema': u'SGA - UNEMI',
                                                                                 'mensaje': mensaje,
                                                                                 'fecha': datetime.now().date(),
                                                                                 'hora': datetime.now().time(),
                                                                                 'saludo': 'Estimada' if programas.inscripcion.persona.sexo_id == 1 else 'Estimado',
                                                                                 'estudiante': programas.inscripcion.persona.nombre_completo_inverso(),
                                                                                 'sexo' : programas.inscripcion.persona.sexo.id,
                                                                                 'proyecto': programas.proyecto.nombre,
                                                                                 'docente': nombre_docente,
                                                                                 'correodocente': correo_docente,
                                                                                 'autoridad2': '',
                                                                                 't': miinstitucion()
                                                                                 }
                                                                        email = programas.inscripcion.persona.lista_emails_envio()
                                                                        send_html_mail(asunto, template, datos, email, [],cuenta=variable_valor('CUENTAS_CORREOS')[0])
                                                                else:
                                                                    observacion = "Registro repetido"
                                                        else:
                                                            if cupos_diferencia != 0:
                                                                if bandera == False:
                                                                    horasvinculacion = haber_cumplido_horas_creditos_vinculacion(inscripcion.id)
                                                                    if not horasvinculacion:
                                                                        if not ProyectoVinculacionInscripcion.objects.filter(status=True,proyectovinculacion=proyectosinvestigacion,inscripcion=inscripcion,periodo=periodoinscripcion).exists():
                                                                            proyectovinculacioninscripcion = ProyectoVinculacionInscripcion(
                                                                                proyectovinculacion=proyectosinvestigacion,
                                                                                inscripcion=inscripcion,
                                                                                periodo=periodoinscripcion,
                                                                                usuario_modificacion=persona.usuario,
                                                                                fecha_creacion=datetime.now().date(),
                                                                                fecha_modificacion=datetime.now().date(),
                                                                                estado=2
                                                                            )
                                                                            proyectovinculacioninscripcion.save(request)

                                                                            if inscripcion.carrera.modalidad == 3:
                                                                                matricula = inscripcion.matricula_periodo(periodo)
                                                                                if materias_asignada_vinculacion := obtener_materia_asignada_vinculacion_por_nivel_v2(matricula.id):
                                                                                    if pdf := materias_asignada_vinculacion.actacompromisovinculacion:
                                                                                        extradetalle = ExtraProyectoVinculacionInscripcion(proyectoinscripcion=proyectovinculacioninscripcion, actacompromisovinculacion=pdf)
                                                                                        extradetalle.save(request)

                                                                            ## -------------------------  INSCRIPCION ------------------------------
                                                                            personal_docente = ParticipantesMatrices.objects.filter(proyecto=proyectovinculacioninscripcion.proyectovinculacion,profesor_id__gt=0, tipoparticipante_id=1,status=True).first()
                                                                            if personal_docente:
                                                                                nombre_docente = personal_docente.profesor.persona.nombre_completo_inverso()
                                                                                correo_docente = personal_docente.profesor.persona.emailinst
                                                                            else:
                                                                                nombre_docente = 'DOCENTE NO ASIGNADO'
                                                                                correo_docente = '-'
                                                                            if not ParticipantesMatrices.objects.filter(proyecto=proyectovinculacioninscripcion.proyectovinculacion,inscripcion=proyectovinculacioninscripcion.inscripcion,status=True).exists():
                                                                                programas = ParticipantesMatrices(matrizevidencia_id=2,
                                                                                                                  proyecto=proyectovinculacioninscripcion.proyectovinculacion,
                                                                                                                  inscripcion=proyectovinculacioninscripcion.inscripcion,
                                                                                                                  horas=0,
                                                                                                                  preinscripcion=proyectovinculacioninscripcion,
                                                                                                                  )
                                                                                programas.save()
                                                                                observacion = "Asignado"

                                                                                saludo = 'Estimada ' if proyectovinculacioninscripcion.inscripcion.persona.sexo_id == 1 else 'Estimado '
                                                                                notificacion = Notificacion(
                                                                                    titulo=f"Asignación de Prácticas de Servicio Comunitario",
                                                                                    cuerpo=f"{saludo}  {proyectovinculacioninscripcion.inscripcion.persona.nombre_completo_inverso()}, nos complace informarte que, has sido asignado para llevar a cabo tus prácticas de servicio comunitario en el proyecto: {programas.proyecto.nombre}, liderado por el docente {nombre_docente}. Por favor, comunícate con el docente responsable, {nombre_docente} al siguiente correo {correo_docente}, para coordinar los detalles específicos de los horarios y actividades a desarrollar dentro de este proceso.",
                                                                                    destinatario=proyectovinculacioninscripcion.inscripcion.persona,
                                                                                    url="/alu_proyectovinculacion?panel=3",
                                                                                    fecha_hora_visible=datetime.now() + timedelta(
                                                                                        days=2),
                                                                                    content_type=None,
                                                                                    object_id=None,
                                                                                    prioridad=1,
                                                                                    app_label='sga')
                                                                                notificacion.save()

                                                                                asunto = "Asignación de Prácticas de Servicio Comunitario"
                                                                                mensaje = "ha sido revisado su informe"
                                                                                template = "emails/aceptacion_preinscripcion_vinc_masiva.html"
                                                                                datos = {'sistema': u'SGA - UNEMI',
                                                                                         'mensaje': mensaje,
                                                                                         'fecha': datetime.now().date(),
                                                                                         'hora': datetime.now().time(),
                                                                                         'saludo': 'Estimada' if programas.inscripcion.persona.sexo_id == 1 else 'Estimado',
                                                                                         'estudiante': programas.inscripcion.persona.nombre_completo_inverso(),
                                                                                         'sexo': programas.inscripcion.persona.sexo.id,
                                                                                         'proyecto': programas.proyecto.nombre,
                                                                                         'docente': nombre_docente,
                                                                                         'correodocente': correo_docente,
                                                                                         'autoridad2': '',
                                                                                         't': miinstitucion()
                                                                                         }
                                                                                email = programas.inscripcion.persona.lista_emails_envio()
                                                                                send_html_mail(asunto, template, datos, email,[], cuenta=variable_valor('CUENTAS_CORREOS')[0])
                                                                        else:
                                                                            observacion = "Registro repetido"

                                                                    else:
                                                                        observacion = "Horas de vinculación completas"

                                                            else:
                                                                observacion = "Cupo no disponible"
                                                    else:
                                                        observacion = 'El estudiante no pertenece a las carreras registradas en el periodo del proyecto'
                                                else:
                                                    observacion = 'El estudiante no pertenece a las carreras registradas en el proyecto'
                                    else:
                                        observacion = "El estudiante no pertenece a las carreras registradas en el proyecto"
                                else:
                                    observacion = "No se encuentra matriculado en el periodo actual"
                                ultima_celda_primera_fila = hoja.cell(row=contador, column=ultima_columna)
                                siguiente_celda = hoja.cell(row=contador, column=ultima_columna + 1)
                                siguiente_celda.font = copy(ultima_celda_primera_fila.font)
                                siguiente_celda.fill = copy(ultima_celda_primera_fila.fill)
                                siguiente_celda.alignment = alin(horizontal="center", vertical="center")
                                siguiente_celda.border = borde
                                siguiente_celda.value = observacion.upper()
                        workbook.save(directory)
                    log(u'%s agregó inscripciónes masivas en el proyecto de vinculación: %s' % (persona, proyectosinvestigacion), request, "add")
                    url =  "/media/innovacion/{}".format(nombre_archivo)
                    notificacion = Notificacion(
                        titulo=f"Reporte de asignación masiva - Vinculación",
                        cuerpo=f"Reporte generado exitosamente con las observaciones corrrespondientes.",
                        destinatario=persona,
                        url=url,
                        fecha_hora_visible=datetime.now() + timedelta(days=2),
                        content_type=None,
                        object_id=None,
                        prioridad=1,
                        app_label='sga')
                    notificacion.save()
                    return JsonResponse({"result": False, "recarga":True, "to": url})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": f'Error: {ex}'})
        return JsonResponse({"result": "bad", "mensaje": u"Solicitud Incorrecta."})

    else:
        if 'action' in request.GET:
            data['action'] = action = request.GET['action']

            if action == 'edittecnicoasociado':
                try:
                    tecnicoasociado = TecnicoAsociadoProyectoVinculacion.objects.get(id=request.GET['id'])
                    f = TecnicoAsociadoProyectoVinculacionForm(initial=model_to_dict(tecnicoasociado))
                    del f.fields['persona']
                    data['form2'] = f
                    data['id'] = tecnicoasociado.pk
                    template = get_template('proyectovinculaciondocente/modal/formmodal.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            if action == 'buscarpersona':
                try:
                    q = request.GET['q'].upper().strip()
                    s = q.split(" ")
                    filtro = Q(usuario__isnull=False, status=True)
                    if len(s) == 1:
                        filtro &= ((Q(nombres__icontains=q) | Q(apellido1__icontains=q) | Q(apellido2__icontains=q) | Q(cedula__contains=q)))
                    elif len(s) == 2:
                        filtro &= ((Q(apellido1__contains=s[0]) & Q(apellido2__contains=s[1])) | (Q(nombres__icontains=s[0]) & Q(nombres__icontains=s[1])) | (Q(nombres__icontains=s[0]) & Q(apellido1__contains=s[1])))
                    else:
                        filtro &= ((Q(nombres__contains=s[0]) & Q(apellido1__contains=s[1]) & Q(apellido2__contains=s[2])) | (Q(nombres__contains=s[0]) & Q(nombres__contains=s[1]) & Q(apellido1__contains=s[2])))

                    per = Persona.objects.filter(filtro).exclude(cedula='').order_by('apellido1', 'apellido2', 'nombres').distinct()[:15]
                    return JsonResponse({"result": "ok", "results": [{"id": x.id, "name": "%s %s" % (f"<img src='{x.get_foto()}' width='25' height='25' style='border-radius: 20%;' alt='...'>", x.nombre_completo_inverso())} for x in per]})
                except Exception as ex:
                    pass

            if action == 'addtecnicoasociado':
                try:
                    f = TecnicoAsociadoProyectoVinculacionForm()
                    data['form2'] = f
                    data['id'] = request.GET.get('id')
                    data['alert'] = {'message': f'Estimad%s <b>%s</b>, de modo que solo puede haber un técnico asociado a este proyecto, al agregar uno nuevo este reemplazará al anterior.' % ('a' if persona.es_mujer() else 'o', persona.nombres.__str__().split(' ')[0])}
                    template = get_template('proyectovinculaciondocente/modal/formmodal.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    transaction.rollback()
                    return JsonResponse({"result": False, 'mensaje': f"Error de conexión. {ex.__str__()}"})

            if action == 'tecnicoasociadoproyectovinculacion':
                try:
                    data['title'] = 'Gestión de técnico asociado'
                    data['proyecto'] = proyecto = ProyectosInvestigacion.objects.get(pk=request.GET.get('id'))
                    data['tecnicos'] = proyecto.get_historialtecnicoasociado().order_by('-activo', '-fechafin')
                    return render(request, 'proyectovinculaciondocente/tecnicoasociadoproyectovinculacion.html', data)
                except Exception as ex:
                    pass

            if action == 'firmainformevinculacion':
                try:
                    data['form2'] = FirmaElectronicaIndividualForm()
                    data['id'] = request.GET.get('id')
                    data['revision'] = request.GET.get('revision', None)
                    data['modal'] = request.GET.get('modal', None)
                    data['action'] = action if not request.GET.get('masivo', None) else 'firmainformevinculacionmasivo'
                    data['pks'] = ', '.join(request.GET.getlist('pks[]'))
                    data['action_to_return'] = 'configurarinforme_adm'
                    template = get_template("proyectovinculaciondocente/modal/firmardocumentoauto.html")
                    return JsonResponse({"result": "ok", 'data': template.render(data)})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

            if action == 'excelprograma':
                    try:
                        __author__ = 'Unemi'

                        output = io.BytesIO()
                        workbook = xlsxwriter.Workbook(output)
                        ws = workbook.add_worksheet('exp_xls_post_part')
                        ws.set_column(0, 0, 10)
                        ws.set_column(1, 1, 40)
                        ws.set_column(2, 2, 50)
                        ws.set_column(3, 3, 20)
                        ws.set_column(4, 4, 20)
                        ws.set_column(5, 5, 20)
                        ws.set_column(6, 6, 40)
                        ws.set_column(7, 7, 35)
                        ws.set_column(8, 8, 40)
                        ws.set_column(9, 9, 20)
                        ws.set_column(10, 10, 20)

                        #                   ws.columm_dimensions['A'].width = 20

                        # formatotitulo = workbook.add_format(
                        #     {'bold': 1, 'text_wrap': True, 'border': 1, 'align': 'center', 'valign': 'middle',
                        #      'fg_color': '#A2D0EC'})
                        formatotitulo_filtros = workbook.add_format(
                            {'bold': 1, 'text_wrap': True, 'border': 1, 'fg_color': '#EBF5FB'})

                        formatoceldacab = workbook.add_format(
                            {'align': 'center', 'bold': 1, 'border': 1, 'text_wrap': True, 'fg_color': '#0b2f44', 'font_color':'white'})
                        formatoceldaleft = workbook.add_format(
                            {'text_wrap': True, 'align': 'center', 'valign': 'vcenter', 'border': 1})

                        ws.write(0, 0, 'COD', formatoceldacab)
                        ws.write(0, 1, 'PROGRAMA', formatoceldacab)
                        ws.write(0, 2, 'LINEA DE INVESTIGACION', formatoceldacab)
                        ws.write(0, 3, 'FECHA INICIO', formatoceldacab)
                        ws.write(0, 4, 'FECHA PLANEADA', formatoceldacab)
                        ws.write(0, 5, 'FECHA REAL', formatoceldacab)
                        ws.write(0, 6, 'AREA DE CONOCIMIENTO', formatoceldacab)
                        ws.write(0, 7, 'SUBAREA DE CONOCIMIENTO', formatoceldacab)
                        ws.write(0, 8, 'SUBAREA ESPECIFICA DE CONOCIMIENTO', formatoceldacab)
                        ws.write(0, 9, 'ALCANCE TERRITORIAL', formatoceldacab)
                        ws.write(0, 10, 'ESTADO', formatoceldacab)

                        listaprogramas = ProgramasInvestigacion.objects.select_related('lineainvestigacion', 'alcanceterritorial',
                                                                                       'areaconocimiento', 'subareaconocimiento',
                                                                                       'subareaespecificaconocimiento').filter(status=True).order_by('-id')

                        filas_recorridas = 2
                        for programa in listaprogramas:
                            est = ""
                            if programa.estado == True:
                                est = "ACTIVO"
                            else:
                                est = "INACTIVO"

                            ws.write('A%s' % filas_recorridas, str(programa.id), formatoceldaleft)
                            ws.write('B%s' % filas_recorridas, str(programa.nombre), formatoceldaleft)
                            ws.write('C%s' % filas_recorridas, str(programa.lineainvestigacion.nombre if programa.lineainvestigacion else "NO REGISTRA"), formatoceldaleft)
                            ws.write('D%s' % filas_recorridas, str(programa.fechainicio), formatoceldaleft)
                            ws.write('E%s' % filas_recorridas, str(programa.fechaplaneado), formatoceldaleft)
                            ws.write('F%s' % filas_recorridas, str(programa.fechareal), formatoceldaleft)
                            ws.write('G%s' % filas_recorridas, str(programa.areaconocimiento.nombre if programa.areaconocimiento else "NO REGISTRA"), formatoceldaleft)
                            ws.write('H%s' % filas_recorridas, str(programa.subareaconocimiento.nombre if programa.subareaconocimiento else "NO REGISTRA"), formatoceldaleft)
                            ws.write('I%s' % filas_recorridas, str(programa.subareaespecificaconocimiento.nombre if programa.subareaespecificaconocimiento else "NO REGISTRA"), formatoceldaleft)
                            ws.write('J%s' % filas_recorridas, str(programa.alcanceterritorial.nombre if programa.alcanceterritorial else "NO REGISTRA"), formatoceldaleft)
                            ws.write('K%s' % filas_recorridas, str(est), formatoceldaleft)

                            filas_recorridas += 1

                        workbook.close()
                        output.seek(0)
                        filename = 'Matriz_Programas.xlsx'
                        response = HttpResponse(output,
                                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                        response['Content-Disposition'] = 'attachment; filename=%s' % filename
                        return response
                    except Exception as ex:
                       pass

            # if action == 'excelprograma':
            #     try:
            #         __author__ = 'Unemi'
            #         style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
            #         style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
            #         style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
            #         title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
            #         style1 = easyxf(num_format_str='D-MMM-YY')
            #         font_style = XFStyle()
            #         font_style.font.bold = True
            #         font_style2 = XFStyle()
            #         font_style2.font.bold = False
            #         wb = Workbook(encoding='utf-8')
            #         ws = wb.add_sheet('exp_xls_post_part')
            #         ws.write_merge(0, 0, 0, 10, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
            #         response = HttpResponse(content_type="application/ms-excel")
            #         response[
            #             'Content-Disposition'] = 'attachment; filename=Matriz_Programas' + random.randint(1, 10000).__str__() + '.xls'
            #
            #         columns = [
            #             (u"NUM", 2000),
            #             (u"NOMBRE", 10000),
            #             (u"LINEA", 10000),
            #             (u"FECHA INICIO", 3000),
            #             (u"FECHA PLANEADA", 3000),
            #             (u"FECHA REAL", 3000),
            #             (u"AREA CONOCIMIENTO", 10000),
            #             (u"SUBAREA CONOCIMIENTO", 10000),
            #             (u"SUBAREA ESPECIFICA", 10000),
            #             (u"ALCANCE TERRITORIAL", 3000),
            #         ]
            #         row_num = 3
            #         for col_num in range(len(columns)):
            #             ws.write(row_num, col_num, columns[col_num][0], font_style)
            #             ws.col(col_num).width = columns[col_num][1]
            #         date_format = xlwt.XFStyle()
            #         date_format.num_format_str = 'yyyy/mm/dd'
            #         listaprogramas = ProgramasInvestigacion.objects.select_related('lineainvestigacion', 'alcanceterritorial',
            #                                                                        'areaconocimiento', 'subareaconocimiento',
            #                                                                        'subareaespecificaconocimiento').filter(status=True).order_by('nombre')
            #         row_num = 4
            #         for programa in listaprogramas:
            #             i = 0
            #             campo1 = programa.id
            #             campo2 = programa.nombre
            #             campo3 = programa.lineainvestigacion.nombre if programa.lineainvestigacion else "NO REGISTRA"
            #             campo4 = programa.fechainicio
            #             campo5 = programa.fechaplaneado
            #             campo6 = programa.fechareal
            #             campo7 = programa.areaconocimiento.nombre if programa.areaconocimiento else "NO REGISTRA"
            #             campo8  = programa.subareaconocimiento.nombre if programa.subareaconocimiento else "NO REGISTRA"
            #             campo9 = programa.subareaespecificaconocimiento.nombre if programa.subareaespecificaconocimiento else "NO REGISTRA"
            #             campo10 = programa.alcanceterritorial.nombre if programa.alcanceterritorial else "NO REGISTRA"
            #             ws.write(row_num, 0, campo1, font_style2)
            #             ws.write(row_num, 1, campo2, font_style2)
            #             ws.write(row_num, 2, campo3, font_style2)
            #             ws.write(row_num, 3, campo4, date_format)
            #             ws.write(row_num, 4, campo5, date_format)
            #             ws.write(row_num, 5, campo6, date_format)
            #             ws.write(row_num, 6, campo7, font_style2)
            #             ws.write(row_num, 7, campo8, font_style2)
            #             ws.write(row_num, 8, campo9, font_style2)
            #             ws.write(row_num, 9, campo10, font_style2)
            #             row_num += 1
            #         wb.save(response)
            #         return response
            #     except Exception as ex:
            #         pass

            elif action == 'excelparticipanteproyecto':
                try:
                    idproyecto = request.GET['idproyecto']
                    __author__ = 'Unemi'
                    style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
                    style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
                    style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
                    title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                    style1 = easyxf(num_format_str='D-MMM-YY')
                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('exp_xls_post_part')
                    ws.write_merge(0, 0, 0, 10, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                    response = HttpResponse(content_type="application/ms-excel")
                    response[
                        'Content-Disposition'] = 'attachment; filename=Matriz_Participantes_Proyectos' + random.randint(
                        1, 10000).__str__() + '.xls'

                    columns = [
                        (u"NUM", 2000),
                        (u"PROGRAMA", 10000),
                        (u"PROYECTO", 10000),
                        (u"CEDULA", 10000),
                        (u"APELLIDOS", 10000),
                        (u"CORREO INSTITUCIONAL", 10000),
                        (u"CORREO PERSONAL", 10000),
                        (u"HORAS", 2000),
                        (u"TIPO", 2000),
                        (u"CARRERA", 10000),
                        (u"ESTADO", 2000),
                        (u"INF. SOLICITADOS", 2000),
                        (u"INF. CARGADOS", 2000),
                        (u"INF. APROBADOS", 2000),
                        (u"INF. RECHAZADOS", 2000),
                        (u"P. INSCRIPCIÓN", 2000),
                    ]
                    row_num = 3
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], font_style)
                        ws.col(col_num).width = columns[col_num][1]
                    date_format = xlwt.XFStyle()
                    date_format.num_format_str = 'yyyy/mm/dd'
                    listaparticipantes = ParticipantesMatrices.objects.filter(status=True, matrizevidencia_id=2, proyecto_id=idproyecto).order_by('inscripcion__persona__apellido1', 'inscripcion__persona__apellido2')
                    row_num = 4
                    for participantes in listaparticipantes:
                        i = 0
                        campo1 = participantes.id
                        campo2 = participantes.proyecto.programa.nombre
                        campo3 = participantes.proyecto.nombre
                        if participantes.profesor:
                            campo4 = participantes.profesor.persona.cedula
                        if participantes.inscripcion:
                            campo4 = participantes.inscripcion.persona.cedula
                            campo8 = participantes.inscripcion.persona.emailinst
                            campo9 = participantes.inscripcion.persona.email
                        if participantes.administrativo:
                            campo4 = participantes.administrativo.persona.cedula
                            campo8 = participantes.administrativo.persona.emailinst
                            campo9 = participantes.administrativo.persona.email
                        if participantes.profesor:
                            campo5 = participantes.profesor.persona.nombre_completo_inverso()
                            campo8 = participantes.profesor.persona.emailinst
                            campo9 = participantes.profesor.persona.email
                        if participantes.inscripcion:
                            campo5 = participantes.inscripcion.persona.nombre_completo_inverso()
                        if participantes.administrativo:
                            campo5 = participantes.administrativo.persona.nombre_completo_inverso()
                        campo6 = participantes.horas
                        if participantes.profesor:
                            campo7 = participantes.tipoparticipante.nombre
                        if participantes.inscripcion:
                            campo7 = 'ESTUDIANTE'
                        if participantes.administrativo:
                            campo7 = 'ADMINISTRATIVO'
                        ws.write(row_num, 0, campo1, font_style2)
                        ws.write(row_num, 1, campo2, font_style2)
                        ws.write(row_num, 2, campo3, font_style2)
                        ws.write(row_num, 3, campo4, font_style2)
                        ws.write(row_num, 4, campo5, font_style2)
                        ws.write(row_num, 5, campo8 , font_style2)
                        ws.write(row_num, 6, campo9 , font_style2)
                        ws.write(row_num, 7, campo6, font_style2)
                        ws.write(row_num, 8, campo7, font_style2)
                        ws.write(row_num, 9, participantes.inscripcion.carrera.nombre + ' ' + participantes.inscripcion.carrera.alias if participantes.inscripcion else "", font_style2)
                        ws.write(row_num, 10, participantes.get_estado_display() if participantes.inscripcion else "", font_style2)
                        ws.write(row_num, 11, InformesProyectoVinculacionDocente.objects.filter(status=True, proyecto_id=idproyecto).count(), font_style2)
                        ws.write(row_num, 12, participantes.cantidad_informes_cargados(), font_style2)
                        ws.write(row_num, 13, participantes.cantidad_informes_aprobados(), font_style2)
                        ws.write(row_num, 14, participantes.cantidad_informes_rechazados(), font_style2)
                        ws.write(row_num, 15, participantes.preinscripcion.periodo.observacion if participantes.preinscripcion else 'Sin periodo' , font_style2)
                        row_num += 1
                    wb.save(response)
                    return response
                except Exception as ex:
                    pass

            elif action == 'excelparticipanteproyectototal':
                try:
                    tipoproyectos = request.GET['tipo']
                    __author__ = 'Unemi'
                    style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
                    style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
                    style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
                    title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                    style1 = easyxf(num_format_str='D-MMM-YY')
                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('exp_xls_post_part')
                    ws.write_merge(0, 0, 0, 10, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                    response = HttpResponse(content_type="application/ms-excel")
                    response[
                        'Content-Disposition'] = 'attachment; filename=Matriz_Participantes_Proyectos' + random.randint(
                        1, 10000).__str__() + '.xls'

                    columns = [
                        (u"NUM", 2000),
                        (u"PROGRAMA", 10000),
                        (u"PROYECTO", 10000),
                        (u"CEDULA", 10000),
                        (u"APELLIDOS", 10000),
                        (u"HORAS", 2000),
                        (u"TIPO", 2000),
                    ]
                    row_num = 3
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], font_style)
                        ws.col(col_num).width = columns[col_num][1]
                    date_format = xlwt.XFStyle()
                    date_format.num_format_str = 'yyyy/mm/dd'
                    listaparticipantes = ParticipantesMatrices.objects.filter(status=True, matrizevidencia_id=2, proyecto__tipo=tipoproyectos)
                    row_num = 4
                    for participantes in listaparticipantes:
                        i = 0
                        campo1 = participantes.id
                        campo2 = participantes.proyecto.programa.nombre
                        campo3 = participantes.proyecto.nombre
                        if participantes.profesor:
                            campo4 = participantes.profesor.persona.cedula
                        if participantes.inscripcion:
                            campo4 = participantes.inscripcion.persona.cedula
                        if participantes.administrativo:
                            campo4 = participantes.administrativo.persona.cedula
                        if participantes.profesor:
                            campo5 = participantes.profesor.persona.nombre_completo()
                        if participantes.inscripcion:
                            campo5 = participantes.inscripcion.persona.nombre_completo()
                        if participantes.administrativo:
                            campo5 = participantes.administrativo.persona.nombre_completo()
                        campo6 = participantes.horas
                        if participantes.profesor:
                            campo7 = participantes.tipoparticipante.nombre
                        if participantes.inscripcion:
                            campo7 = 'ESTUDIANTE'
                        if participantes.administrativo:
                            campo7 = 'ADMINISTRATIVO'
                        ws.write(row_num, 0, campo1, font_style2)
                        ws.write(row_num, 1, campo2, font_style2)
                        ws.write(row_num, 2, campo3, font_style2)
                        ws.write(row_num, 3, campo4, font_style2)
                        ws.write(row_num, 4, campo5, font_style2)
                        ws.write(row_num, 5, campo6, font_style2)
                        ws.write(row_num, 6, campo7, font_style2)
                        row_num += 1
                    wb.save(response)
                    return response
                except Exception as ex:
                    pass

            # elif action == 'excelparticipantevinculacion':
            #     try:
            #         tipoproyectos = request.GET['tipo']
            #         __author__ = 'Unemi'
            #         style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
            #         style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
            #         style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
            #         title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
            #         style1 = easyxf(num_format_str='D-MMM-YY')
            #         font_style = XFStyle()
            #         font_style.font.bold = True
            #         font_style2 = XFStyle()
            #         font_style2.font.bold = False
            #         wb = Workbook(encoding='utf-8')
            #         ws = wb.add_sheet('exp_xls_post_part')
            #         ws.write_merge(0, 0, 0, 10, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
            #         response = HttpResponse(content_type="application/ms-excel")
            #         response[
            #             'Content-Disposition'] = 'attachment; filename=Matriz_Participantes_Proyectos' + random.randint(
            #             1, 10000).__str__() + '.xls'
            #
            #         columns = [
            #             (u"NUM", 2000),
            #             (u"PROGRAMA", 10000),
            #             (u"PROYECTO", 10000),
            #             (u"ANIO", 2000),
            #             (u"CARRERAS PROYECTO", 10000),
            #             (u"CEDULA", 3000),
            #             (u"APELLIDOS", 10000),
            #             (u"CARRERAS ESTUDIANTE", 10000),
            #             (u"HORAS", 2000),
            #             (u"TIPO", 2000),
            #         ]
            #         row_num = 3
            #         for col_num in range(len(columns)):
            #             ws.write(row_num, col_num, columns[col_num][0], font_style)
            #             ws.col(col_num).width = columns[col_num][1]
            #         date_format = xlwt.XFStyle()
            #         date_format.num_format_str = 'yyyy/mm/dd'
            #         listaparticipantes = ParticipantesMatrices.objects.filter(status=True, matrizevidencia_id=2, proyecto__tipo=tipoproyectos).order_by('inscripcion__persona__apellido1','inscripcion__persona__apellido2')
            #         row_num = 4
            #         for participantes in listaparticipantes:
            #             lista = []
            #             i = 0
            #             campo1 = participantes.id
            #             campo2 = participantes.proyecto.programa.nombre
            #             campo3 = participantes.proyecto.nombre
            #             campocarrera = ''
            #             anio = str(participantes.proyecto.fechainicio)
            #             anio = anio.split('-')
            #             if participantes.profesor:
            #                 campo4 = participantes.profesor.persona.cedula
            #             if participantes.inscripcion:
            #                 campo4 = participantes.inscripcion.persona.cedula
            #                 campocarrera = participantes.inscripcion.carrera.nombre + ' ' + participantes.inscripcion.carrera.alias
            #             if participantes.administrativo:
            #                 campo4 = participantes.administrativo.persona.cedula
            #             if participantes.profesor:
            #                 campo5 = participantes.profesor.persona.nombre_completo_inverso()
            #             if participantes.inscripcion:
            #                 campo5 = participantes.inscripcion.persona.nombre_completo_inverso()
            #             if participantes.administrativo:
            #                 campo5 = participantes.administrativo.persona.nombre_completo_inverso()
            #             campo6 = participantes.horas
            #             if participantes.profesor:
            #                 campo7 = participantes.tipoparticipante.nombre
            #             if participantes.inscripcion:
            #                 campo7 = 'ESTUDIANTE'
            #             if participantes.administrativo:
            #                 campo7 = 'ADMINISTRATIVO'
            #
            #             ws.write(row_num, 0, campo1, font_style2)
            #             ws.write(row_num, 1, campo2, font_style2)
            #             ws.write(row_num, 2, campo3, font_style2)
            #             if participantes.inscripcion:
            #                 if participantes.proyecto.carrerasproyecto_set.filter(status=True).exists():
            #                     carrerasproyectos = participantes.proyecto.carrerasproyecto_set.filter(status=True)
            #                     for carrerasproy in carrerasproyectos:
            #                         lista.append(carrerasproy.carrera.nombre + ',')
            #             ws.write(row_num, 3, anio[0], font_style2)
            #             ws.write(row_num, 4, str(lista), font_style2)
            #             ws.write(row_num, 5, campo4, font_style2)
            #             ws.write(row_num, 6, campo5, font_style2)
            #             ws.write(row_num, 7, campocarrera, font_style2)
            #             ws.write(row_num, 8, campo6, font_style2)
            #             ws.write(row_num, 9, campo7, font_style2)
            #             row_num += 1
            #         wb.save(response)
            #         return response
            #     except Exception as ex:
            #         pass

            elif action == 'excelparticipantevinculacion':
                    try:
                        __author__ = 'Unemi'

                        tipoproyectos = request.GET['tipo']

                        output = io.BytesIO()
                        workbook = xlsxwriter.Workbook(output)
                        ws = workbook.add_worksheet('exp_xls_post_part')
                        ws.set_column(0, 0, 10)
                        ws.set_column(1, 1, 40)
                        ws.set_column(2, 2, 60)
                        ws.set_column(3, 3, 10)
                        ws.set_column(4, 4, 40)
                        ws.set_column(5, 5, 15)
                        ws.set_column(6, 6, 40)
                        ws.set_column(7, 7, 40)
                        ws.set_column(8, 8, 10)
                        ws.set_column(9, 9, 20)

                        #                   ws.columm_dimensions['A'].width = 20

                        # formatotitulo = workbook.add_format(
                        #     {'bold': 1, 'text_wrap': True, 'border': 1, 'align': 'center', 'valign': 'middle',
                        #      'fg_color': '#A2D0EC'})
                        formatotitulo_filtros = workbook.add_format(
                            {'bold': 1, 'text_wrap': True, 'border': 1, 'fg_color': '#EBF5FB'})

                        formatoceldacab = workbook.add_format(
                            {'align': 'center', 'bold': 1, 'border': 1, 'text_wrap': True, 'fg_color': '#0b2f44', 'font_color':'white'})
                        formatoceldaleft = workbook.add_format(
                            {'text_wrap': True, 'align': 'center', 'valign': 'vcenter', 'border': 1})

                        ws.write(0, 0, 'COD', formatoceldacab)
                        ws.write(0, 1, 'PROGRAMA', formatoceldacab)
                        ws.write(0, 2, 'PROYECTO', formatoceldacab)
                        ws.write(0, 3, 'ANIO', formatoceldacab)
                        ws.write(0, 4, 'CARRERAS PROYECTO', formatoceldacab)
                        ws.write(0, 5, 'CEDULA', formatoceldacab)
                        ws.write(0, 6, 'PARTICIPANTES', formatoceldacab)
                        ws.write(0, 7, 'CARRERAS ESTUDIANTE', formatoceldacab)
                        ws.write(0, 8, 'HORAS', formatoceldacab)
                        ws.write(0, 9, 'TIPO', formatoceldacab)

                        listaparticipantes = ParticipantesMatrices.objects.filter(status=True, matrizevidencia_id=2, proyecto__tipo=tipoproyectos).order_by('inscripcion__persona__apellido1',
                                                                                                                                                            'inscripcion__persona__apellido2')

                        filas_recorridas = 2
                        for participantes in listaparticipantes:
                            lista = []
                            i = 0
                            anio = str(participantes.proyecto.fechainicio)
                            anio = anio.split('-')
                            campocarrera = ''

                            if participantes.profesor:
                                campo4 = participantes.profesor.persona.cedula
                            if participantes.inscripcion:
                                campo4 = participantes.inscripcion.persona.cedula
                                campocarrera = participantes.inscripcion.carrera.nombre + ' ' + participantes.inscripcion.carrera.alias
                            if participantes.administrativo:
                                campo4 = participantes.administrativo.persona.cedula
                            if participantes.profesor:
                                campo5 = participantes.profesor.persona.nombre_completo_inverso()
                            if participantes.inscripcion:
                                campo5 = participantes.inscripcion.persona.nombre_completo_inverso()
                            if participantes.administrativo:
                                campo5 = participantes.administrativo.persona.nombre_completo_inverso()
                            if participantes.profesor:
                                campo7 = participantes.tipoparticipante.nombre
                            if participantes.inscripcion:
                                campo7 = 'ESTUDIANTE'
                            if participantes.administrativo:
                                campo7 = 'ADMINISTRATIVO'

                            ws.write('A%s' % filas_recorridas, str(participantes.id), formatoceldaleft)
                            ws.write('B%s' % filas_recorridas, str(participantes.proyecto.programa.nombre), formatoceldaleft)
                            ws.write('C%s' % filas_recorridas, str(participantes.proyecto.nombre), formatoceldaleft)
                            if participantes.inscripcion:
                                if participantes.proyecto.carrerasproyecto_set.filter(status=True).exists():
                                    carrerasproyectos = participantes.proyecto.carrerasproyecto_set.filter(status=True)
                                    for carrerasproy in carrerasproyectos:
                                        lista.append(carrerasproy.carrera.nombre + ',')
                            ws.write('D%s' % filas_recorridas, str(anio[0]), formatoceldaleft)
                            ws.write('E%s' % filas_recorridas, str(lista if lista else 'NO REGISTRA'), formatoceldaleft)
                            ws.write('F%s' % filas_recorridas, str(campo4), formatoceldaleft)
                            ws.write('G%s' % filas_recorridas, str(campo5), formatoceldaleft)
                            ws.write('H%s' % filas_recorridas, str(campocarrera if campocarrera else 'NO REGISTRA'), formatoceldaleft)
                            ws.write('I%s' % filas_recorridas, str(participantes.horas), formatoceldaleft)
                            ws.write('J%s' % filas_recorridas, str(campo7), formatoceldaleft)

                            filas_recorridas += 1

                        workbook.close()
                        output.seek(0)
                        filename = 'Matriz_Participantes_Proyectos.xlsx'
                        response = HttpResponse(output,
                                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                        response['Content-Disposition'] = 'attachment; filename=%s' % filename
                        return response
                    except Exception as ex:
                       pass


            elif action == 'excelmatriculadosvinculacion':
                    try:
                        __author__ = 'Unemi'

                        output = io.BytesIO()
                        workbook = xlsxwriter.Workbook(output)
                        ws = workbook.add_worksheet('exp_xls_post_part')
                        ws.set_column(0, 0, 10)
                        ws.set_column(1, 1, 40)
                        ws.set_column(2, 2, 40)
                        ws.set_column(3, 3, 20)
                        ws.set_column(4, 4, 40)
                        ws.set_column(5, 5, 20)

                        #                   ws.columm_dimensions['A'].width = 20

                        # formatotitulo = workbook.add_format(
                        #     {'bold': 1, 'text_wrap': True, 'border': 1, 'align': 'center', 'valign': 'middle',
                        #      'fg_color': '#A2D0EC'})
                        formatotitulo_filtros = workbook.add_format(
                            {'bold': 1, 'text_wrap': True, 'border': 1, 'fg_color': '#EBF5FB'})

                        formatoceldacab = workbook.add_format(
                            {'align': 'center', 'bold': 1, 'border': 1, 'text_wrap': True, 'fg_color': '#0b2f44', 'font_color':'white'})
                        formatoceldaleft = workbook.add_format(
                            {'text_wrap': True, 'align': 'center', 'valign': 'vcenter', 'border': 1})

                        ws.write(0, 0, 'COD', formatoceldacab)
                        ws.write(0, 1, 'FACULTAD', formatoceldacab)
                        ws.write(0, 2, 'CARRERA', formatoceldacab)
                        ws.write(0, 3, 'CEDULA', formatoceldacab)
                        ws.write(0, 4, 'NOMBRES', formatoceldacab)
                        ws.write(0, 5, 'HORAS', formatoceldacab)

                        cursor = connection.cursor()
                        listaestudiante = "select coor.nombre as nomfacultad, " \
                                          "coor.alias as nomalias,carr.nombre as carrera,carr.alias as aliascarrera,per.cedula,per.apellido1, " \
                                          "per.apellido2,per.nombres, " \
                                          "(select sum(matrices.horas) " \
                                          "from sga_participantesmatrices matrices,sga_proyectosinvestigacion proy " \
                                          "where matrices.proyecto_id=proy.id " \
                                          "and matrices.matrizevidencia_id=2 " \
                                          "and matrices.inscripcion_id=ins.id " \
                                          "and matrices.status=true " \
                                          "and proy.status=true " \
                                          "and proy.tipo=1) as horasvinculacion " \
                                          "from sga_matricula matri,sga_inscripcion ins,sga_coordinacion coor,sga_nivel ni, " \
                                          "sga_persona per,sga_carrera carr " \
                                          "where ins.coordinacion_id=coor.id " \
                                          "and ins.carrera_id=carr.id " \
                                          "and ins.persona_id=per.id " \
                                          "and matri.inscripcion_id=ins.id " \
                                          "and matri.retiradomatricula=false " \
                                          "and matri.status=true " \
                                          "and matri.nivel_id=ni.id " \
                                          "and ni.periodo_id=" + str(periodo.id) + " " \
                                                                                   "and coor.id not in (9) " \
                                                                                   "and matri.estado_matricula in (2,3) " \
                                                                                   "and matri.id not in (select tabla.matricula_id from " \
                                                                                   "(select count(matri.id) as contar, matri.id as matricula_id " \
                                                                                   "from sga_matricula matri,sga_inscripcion ins,sga_coordinacion coor,sga_nivel ni, " \
                                                                                   "sga_materiaasignada matasig,sga_materia mate, sga_asignatura asig,sga_carrera carr " \
                                                                                   "where ins.coordinacion_id=coor.id " \
                                                                                   "and ins.carrera_id=carr.id " \
                                                                                   "and matri.inscripcion_id=ins.id " \
                                                                                   "and matri.retiradomatricula=false " \
                                                                                   "and matri.status=true " \
                                                                                   "and matri.nivel_id=ni.id " \
                                                                                   "and carr.id not in (7) " \
                                                                                   "and matasig.matricula_id=matri.id " \
                                                                                   "and matasig.materia_id=mate.id " \
                                                                                   "and mate.asignatura_id=asig.id " \
                                                                                   "and ni.periodo_id=" + str(periodo.id) + " " \
                                                                                                                            "and matri.estado_matricula in (2,3) " \
                                                                                                                            "GROUP by matri.id) as tabla, " \
                                                                                                                            "sga_materiaasignada matasig1,sga_materia mate1, sga_asignatura asig1 " \
                                                                                                                            "where tabla.contar = 1 and matasig1.matricula_id=tabla.matricula_id " \
                                                                                                                            "and matasig1.materia_id=mate1.id and mate1.asignatura_id=asig1.id " \
                                                                                                                            "and asig1.modulo=True) " \
                                                                                                                            "order by per.apellido1,per.apellido2,per.nombres"

                        cursor.execute(listaestudiante)
                        results = cursor.fetchall()

                        filas_recorridas = 2
                        i = 0
                        for result in results:
                            i += 1
                            ws.write('A%s' % filas_recorridas, str(i), formatoceldaleft)
                            ws.write('B%s' % filas_recorridas, str(result[0]), formatoceldaleft)
                            ws.write('C%s' % filas_recorridas, str(result[2]), formatoceldaleft)
                            ws.write('D%s' % filas_recorridas, str(result[4]), formatoceldaleft)
                            ws.write('E%s' % filas_recorridas, str(result[5] + ' ' + result[6] + ' ' + result[7]), formatoceldaleft)
                            ws.write('F%s' % filas_recorridas, str(result[8] if result[8] else "NO REGISTRA"), formatoceldaleft)

                            filas_recorridas += 1

                        workbook.close()
                        output.seek(0)
                        filename = 'Matriculados_Vinculacion.xlsx'
                        response = HttpResponse(output,
                                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                        response['Content-Disposition'] = 'attachment; filename=%s' % filename
                        return response
                    except Exception as ex:
                       pass

            # elif action == 'excelmatriculadosvinculacion':
            #     try:
            #         tipoproyectos = request.GET['tipo']
            #         __author__ = 'Unemi'
            #         style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
            #         style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
            #         style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
            #         title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
            #         style1 = easyxf(num_format_str='D-MMM-YY')
            #         font_style = XFStyle()
            #         font_style.font.bold = True
            #         font_style2 = XFStyle()
            #         font_style2.font.bold = False
            #         wb = Workbook(encoding='utf-8')
            #         ws = wb.add_sheet('exp_xls_post_part')
            #         ws.write_merge(0, 0, 0, 10, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
            #         response = HttpResponse(content_type="application/ms-excel")
            #         response[
            #             'Content-Disposition'] = 'attachment; filename=Matriculados_Vinculacion' + random.randint(
            #             1, 10000).__str__() + '.xls'
            #
            #         columns = [
            #             (u"NUM", 2000),
            #             (u"FACULTAD", 10000),
            #             (u"CARRERA", 15000),
            #             (u"CEDULA", 3000),
            #             (u"NOMBRES", 15000),
            #             (u"HORAS", 2000),
            #         ]
            #         row_num = 3
            #         for col_num in range(len(columns)):
            #             ws.write(row_num, col_num, columns[col_num][0], font_style)
            #             ws.col(col_num).width = columns[col_num][1]
            #         date_format = xlwt.XFStyle()
            #         date_format.num_format_str = 'yyyy/mm/dd'
            #         cursor = connection.cursor()
            #         listaestudiante = "select coor.nombre as nomfacultad, " \
            #                             "coor.alias as nomalias,carr.nombre as carrera,carr.alias as aliascarrera,per.cedula,per.apellido1, " \
            #                             "per.apellido2,per.nombres, " \
            #                             "(select sum(matrices.horas) " \
            #                           "from sga_participantesmatrices matrices,sga_proyectosinvestigacion proy " \
            #                           "where matrices.proyecto_id=proy.id " \
            #                           "and matrices.matrizevidencia_id=2 " \
            #                           "and matrices.inscripcion_id=ins.id " \
            #                           "and matrices.status=true " \
            #                            "and proy.status=true " \
            #                             "and proy.tipo=1) as horasvinculacion " \
            #                             "from sga_matricula matri,sga_inscripcion ins,sga_coordinacion coor,sga_nivel ni, " \
            #                             "sga_persona per,sga_carrera carr " \
            #                             "where ins.coordinacion_id=coor.id " \
            #                             "and ins.carrera_id=carr.id " \
            #                             "and ins.persona_id=per.id " \
            #                             "and matri.inscripcion_id=ins.id " \
            #                             "and matri.retiradomatricula=false " \
            #                             "and matri.status=true " \
            #                             "and matri.nivel_id=ni.id " \
            #                             "and ni.periodo_id="+ str(periodo.id) +" " \
            #                             "and coor.id not in (9) " \
            #                             "and matri.estado_matricula in (2,3) " \
            #                             "and matri.id not in (select tabla.matricula_id from " \
            #                             "(select count(matri.id) as contar, matri.id as matricula_id " \
            #                             "from sga_matricula matri,sga_inscripcion ins,sga_coordinacion coor,sga_nivel ni, " \
            #                             "sga_materiaasignada matasig,sga_materia mate, sga_asignatura asig,sga_carrera carr " \
            #                             "where ins.coordinacion_id=coor.id " \
            #                             "and ins.carrera_id=carr.id " \
            #                             "and matri.inscripcion_id=ins.id " \
            #                             "and matri.retiradomatricula=false " \
            #                             "and matri.status=true " \
            #                             "and matri.nivel_id=ni.id " \
            #                             "and carr.id not in (7) " \
            #                             "and matasig.matricula_id=matri.id " \
            #                             "and matasig.materia_id=mate.id " \
            #                             "and mate.asignatura_id=asig.id " \
            #                             "and ni.periodo_id="+ str(periodo.id) +" " \
            #                             "and matri.estado_matricula in (2,3) " \
            #                             "GROUP by matri.id) as tabla, " \
            #                             "sga_materiaasignada matasig1,sga_materia mate1, sga_asignatura asig1 " \
            #                             "where tabla.contar = 1 and matasig1.matricula_id=tabla.matricula_id " \
            #                             "and matasig1.materia_id=mate1.id and mate1.asignatura_id=asig1.id " \
            #                             "and asig1.modulo=True) " \
            #                             "order by per.apellido1,per.apellido2,per.nombres"
            #
            #         cursor.execute(listaestudiante)
            #         results = cursor.fetchall()
            #         row_num = 4
            #         i = 0
            #         for resul in results:
            #             i += 1
            #             ws.write(row_num, 0, i, font_style2)
            #             ws.write(row_num, 1, resul[0], font_style2)
            #             ws.write(row_num, 2, resul[2], font_style2)
            #             ws.write(row_num, 3, resul[4], font_style2)
            #             ws.write(row_num, 4, resul[5] + ' ' + resul[6] + ' ' + resul[7], font_style2)
            #             ws.write(row_num, 5, resul[8], font_style2)
            #             row_num += 1
            #         wb.save(response)
            #         return response
            #     except Exception as ex:
            #         pass

            elif action == 'excelestudiantesinscrito':
                try:
                    idperiodo = request.GET['id']
                    __author__ = 'Unemi'
                    style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
                    style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
                    style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
                    title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                    style1 = easyxf(num_format_str='D-MMM-YY')
                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('exp_xls_post_part')
                    ws.write_merge(0, 0, 0, 10, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                    response = HttpResponse(content_type="application/ms-excel")
                    response[
                        'Content-Disposition'] = 'attachment; filename=Matriz_Participantes_Proyectos' + random.randint(
                        1, 10000).__str__() + '.xls'

                    columns = [
                        (u"NUM", 2000),
                        (u"PROYECTO", 10000),
                        (u"CEDULA", 10000),
                        (u"APELLIDOS / NOMBRES", 10000),
                        (u"CARRERA", 10000),
                        (u"NIVEL MATRICULA", 10000),
                        (u"CORREO PERSONAL", 10000),
                        (u"CORREO INSTITUCIONAL", 10000),
                        (u"CELULAR", 10000),
                        (u"ESTADO", 10000),
                        (u"HORAS REALIZADAS", 10000),
                    ]
                    row_num = 3
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], font_style)
                        ws.col(col_num).width = columns[col_num][1]
                    date_format = xlwt.XFStyle()
                    date_format.num_format_str = 'yyyy/mm/dd'
                    listaparticipantes = ProyectoVinculacionInscripcion.objects.filter(status=True, periodo__id =idperiodo ).order_by('inscripcion__persona__apellido1', 'inscripcion__persona__apellido2')
                    row_num = 4
                    i = 0
                    for participantes in listaparticipantes:
                        i+=1
                        if participantes.estado == 1:
                            estado= "SOLICITADO"
                        elif participantes.estado == 2:
                            estado = "APROBADO"
                        else:
                            estado = "RECHAZADO"
                        matricula = participantes.inscripcion.matricula_periodo(participantes.periodo.periodo)
                        if matricula:
                            nivel = matricula.nivelmalla
                        else:
                            nivel = participantes.inscripcion.mi_nivel().nivel
                        ws.write(row_num, 0, i, font_style2)
                        ws.write(row_num, 1, participantes.proyectovinculacion.nombre, font_style2)
                        ws.write(row_num, 2, participantes.inscripcion.persona.cedula, font_style2)
                        ws.write(row_num, 3, participantes.inscripcion.persona.nombre_completo_inverso(), font_style2)
                        ws.write(row_num, 4, participantes.inscripcion.carrera.nombre, font_style2)
                        ws.write(row_num, 5, str(nivel), font_style2)
                        ws.write(row_num, 6, participantes.inscripcion.persona.email, font_style2)
                        ws.write(row_num, 7, participantes.inscripcion.persona.emailinst, font_style2)
                        ws.write(row_num, 8, participantes.inscripcion.persona.telefono, font_style2)
                        ws.write(row_num, 9, estado, font_style2)
                        ws.write(row_num, 10, participantes.inscripcion.numero_horas_proyectos_vinculacion(), font_style2)
                        row_num += 1
                    wb.save(response)
                    return response
                except Exception as ex:
                    pass


            elif action == 'excelactividadesinforme':
                try:
                    __author__ = 'Unemi'
                    conf = ConfiguracionInformeVinculacion.objects.get(id=request.GET['id'])
                    tareas = DetalleInformeVinculacion.objects.filter(status=True,configuracion__id=(request.GET['id']),tarea__isnull=False)
                    proyecto = ProyectosInvestigacion.objects.get(pk=conf.proyecto.pk)
                    fines = DetalleInformeVinculacion.objects.filter(status=True, configuracion=conf, proyecto=proyecto,actividad__arbolObjetivo__tipo=3)
                    proposito = DetalleInformeVinculacion.objects.filter(status=True, configuracion=conf,proyecto=proyecto,actividad__arbolObjetivo__tipo=1)
                    actividades = MarcoLogicoReporte.objects.filter(status=True, configuracion=conf, proyecto=proyecto,arbolObjetivo__tipo=2,arbolObjetivo__parentID__isnull=False)
                    componentes = MarcoLogicoReporte.objects.filter(status=True, configuracion=conf, proyecto=proyecto,arbolObjetivo__tipo=2,arbolObjetivo__parentID__isnull=True)

                    output = io.BytesIO()
                    workbook = xlsxwriter.Workbook(output)
                    ws = workbook.add_worksheet('Reporte')
                    ws.set_column(0, 0, 25)
                    ws.set_column(1, 1, 25)
                    ws.set_column(2, 2, 25)
                    ws.set_column(3, 3, 10)
                    ws.set_column(4, 4, 13)
                    ws.set_column(5, 5, 13)
                    ws.set_column(6, 6, 40)
                    ws.set_column(7, 7, 25)
                    ws.set_column(8, 8, 15)
                    ws.set_column(9, 9, 15)
                    ws.set_column(10, 9, 15)

                    formatotitulo_filtros = workbook.add_format({'bold': 1, 'text_wrap': True, 'border': 1, 'fg_color': '#EBF5FB'})
                    # formatoceldacab = workbook.add_format({'align': 'center', 'bold': 1, 'border': 1, 'text_wrap': True, 'fg_color': '#0b2f44','font_color': 'white'})
                    formatoceldacab = workbook.add_format({'align': 'center', 'bold': 1,'font_size':14})
                    formatoceldatituloleft = workbook.add_format({'align': 'left', 'bold': 1, 'border': 1, 'text_wrap': True})
                    formatoceldatitulocentro = workbook.add_format({'align': 'center', 'bold': 1, 'border': 1, 'text_wrap': True})
                    formatoceldaleft = workbook.add_format({'text_wrap': True, 'align': 'center', 'valign': 'vcenter', 'border': 1})

                    ws.merge_range('A%s:A%s' % (2, 9), '')
                    ws.merge_range('B%s:H%s' % (2, 2), 'VICERRECTORADO DE VINCULACIÓN',formatoceldacab)
                    ws.merge_range('B%s:H%s' % (3, 3), 'SERVICIO COMUNITARIO',formatoceldacab)
                    ws.merge_range('B%s:H%s' % (4, 4), 'FICHA DE CONTROL DE SEGUIMIENTO Y MONITOREO',formatoceldacab)
                    ws.merge_range('B%s:H%s' % (5, 5), '')
                    ws.merge_range('B%s:G%s' % (6, 6), 'PROYECTO: '+proyecto.nombre, formatoceldaleft)
                    ws.merge_range('H%s:H%s' % (6, 9), 'CÓDIGO: '+str(proyecto.pk), formatoceldaleft)
                    ws.merge_range('B%s:G%s' % (7, 7), 'CARRERA: ', formatoceldaleft)
                    ws.merge_range('B%s:G%s' % (8, 8), 'BENEFICIARIOS: ', formatoceldaleft)
                    ws.merge_range('B%s:G%s' % (9, 9), 'LÍDER: ' + proyecto.lider().nombre_completo_inverso() , formatoceldaleft)
                    ws.write('A%s' % 11, 'INFORME DE AVANCE:')
                    ws.write('E%s' % 11, 'FECHA DEL SEGUIMIENTO:')

                    filas_recorridas = 13
                    for componente in componentes:

                        ws.merge_range('A%s:F%s' % (filas_recorridas, filas_recorridas), 'DESCRIPCIÓN DEL RESULTADO',formatoceldatituloleft)
                        ws.write('G%s' % filas_recorridas, 'ALCANCE', formatoceldatitulocentro)
                        ws.write('H%s' % filas_recorridas, 'AVANCE', formatoceldatitulocentro)
                        filas_recorridas += 1
                        ws.merge_range('A%s:F%s' % (filas_recorridas, filas_recorridas), componente.arbolObjetivo.detalle,formatoceldatituloleft)
                        ws.write('G%s' % filas_recorridas, componente.cumplimiento, formatoceldaleft)
                        ws.write('H%s' % filas_recorridas, componente.avance, formatoceldaleft)
                        ws.write('I%s' % filas_recorridas, componente.avancemensual, formatoceldaleft)

                        filas_recorridas += 1
                        for actividad in actividades:
                            if componente.arbolObjetivo.pk == actividad.arbolObjetivo.parentID.pk:

                                ws.merge_range('A%s:F%s' % (filas_recorridas, filas_recorridas), actividad.arbolObjetivo.detalle,formatoceldatituloleft)
                                ws.write('G%s' % filas_recorridas, actividad.cumplimiento, formatoceldaleft)
                                ws.write('H%s' % filas_recorridas, actividad.avance, formatoceldaleft)
                                ws.write('I%s' % filas_recorridas, actividad.avancemensual, formatoceldaleft)
                                filas_recorridas += 1
                                ws.write('A%s' % filas_recorridas, 'INDICADOR', formatoceldatitulocentro)
                                ws.write('B%s' % filas_recorridas, 'TAREAS', formatoceldatitulocentro)
                                ws.write('C%s' % filas_recorridas, 'EVIDENCIA', formatoceldatitulocentro)
                                ws.write('D%s' % filas_recorridas, 'FECHA INICIO', formatoceldatitulocentro)
                                ws.write('E%s' % filas_recorridas, 'FECHA FIN PLANIFICADA', formatoceldatitulocentro)
                                ws.write('F%s' % filas_recorridas, 'FECHA FIN REAL', formatoceldatitulocentro)
                                ws.write('G%s' % filas_recorridas, 'DOCENTES RESPONSABLES', formatoceldatitulocentro)
                                ws.write('H%s' % filas_recorridas, 'OBSERVACIONES', formatoceldatitulocentro)
                                ws.write('I%s' % filas_recorridas, 'ALCANCE', formatoceldatitulocentro)
                                ws.write('J%s' % filas_recorridas, 'AVANCE AC.', formatoceldatitulocentro)
                                ws.write('K%s' % filas_recorridas, 'AVANCE M.', formatoceldatitulocentro)

                                filas_recorridas += 1
                                i = 0
                                for tarea in tareas:
                                    if tarea.tarea.aobjetivo.pk == actividad.arbolObjetivo.pk:
                                        marco = MarcoLogico.objects.get(status=True, arbolObjetivo=tarea.tarea.aobjetivo)
                                        fin_real = DetalleCumplimiento.objects.filter(tarea=tarea.tarea).latest('fecha_ingreso').fecha_ingreso
                                        responsable = ''
                                        for docen in tarea.tarea.responsable.all():
                                            responsable = responsable + docen.persona.nombre_completo_inverso() + '\n'
                                        ws.write('A%s' % filas_recorridas, marco.indicador, formatoceldaleft)
                                        ws.write('B%s' % filas_recorridas, tarea.tarea.descripcion,formatoceldaleft)
                                        ws.write('C%s' % filas_recorridas, '', formatoceldaleft)
                                        ws.write('D%s' % filas_recorridas, str(tarea.tarea.fecha_inicio), formatoceldaleft)
                                        ws.write('E%s' % filas_recorridas, str(tarea.tarea.fecha_fin), formatoceldaleft)
                                        ws.write('F%s' % filas_recorridas, str(fin_real), formatoceldaleft)
                                        ws.write('G%s' % filas_recorridas, responsable, formatoceldaleft)
                                        ws.write('H%s' % filas_recorridas, '', formatoceldaleft)
                                        ws.write('I%s' % filas_recorridas, tarea.tarea.cumplimiento, formatoceldaleft)
                                        ws.write('J%s' % filas_recorridas, tarea.avanceacumulado, formatoceldaleft)
                                        ws.write('K%s' % filas_recorridas, tarea.porcentaje_avance, formatoceldaleft)
                                        filas_recorridas += 1
                    ws.write('G%s' % filas_recorridas, 'AVANCE DEL PROPÓSITO', formatoceldaleft)
                    ws.write('H%s' % filas_recorridas, conf.avanceproposito().porcentaje_avance, formatoceldaleft)
                    filas_recorridas += 2

                    ws.write('A%s' % filas_recorridas, 'SUGERENCIAS Y/O OBSERVACIONES:', formatoceldaleft)
                    ws.merge_range('B%s:H%s' % (filas_recorridas, filas_recorridas), '',formatoceldaleft)
                    ws.write('A%s' % (filas_recorridas+1), 'Código del formato FR069-V1.00-2022-UNEMI')
                    filas_recorridas += 6
                    ws.merge_range('A%s:H%s' % (filas_recorridas, filas_recorridas), persona.nombre_completo_inverso(), formatoceldaleft)
                    ws.merge_range('A%s:H%s' % ((filas_recorridas+1), (filas_recorridas+1)), 'TÉCNICO DOCENTE DE VINCULACIÓN', formatoceldaleft)
                            # ws.write_formula('G%s' % filas_recorridas,'=SUM(E%s+F%s)' % (filas_recorridas, filas_recorridas), formatoceldaleft)
                    # ws.write('F%s' % filas_recorridas, 'Total', formatoceldaleft)
                    # ws.write_formula('G%s' % filas_recorridas, '=SUM(G2:G%s)' % (filas_recorridas - 1),formatoceldaleft)
                    workbook.close()
                    output.seek(0)

                    filename = 'Reporte.xlsx'
                    response = HttpResponse(output,content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    response['Content-Disposition'] = 'attachment; filename=%s' % filename

                    return response

                except Exception as ex:

                    pass

            elif action == 'excelproyectosinvestigacionvinculacion':
                    try:
                        __author__ = 'Unemi'

                        tipoproyectos = request.GET['tipo']

                        output = io.BytesIO()
                        workbook = xlsxwriter.Workbook(output)
                        ws = workbook.add_worksheet('exp_xls_post_part')
                        ws.set_column(0, 0, 10)
                        ws.set_column(1, 1, 50)
                        ws.set_column(2, 2, 30)
                        ws.set_column(3, 3, 30)
                        ws.set_column(4, 4, 30)
                        ws.set_column(5, 5, 30)
                        ws.set_column(6, 6, 60)
                        ws.set_column(7, 7, 60)
                        ws.set_column(8, 8, 20)
                        ws.set_column(9, 9, 20)
                        ws.set_column(10, 10, 20)
                        ws.set_column(11, 11, 20)
                        ws.set_column(12, 12, 40)
                        ws.set_column(13, 13, 20)
                        ws.set_column(14, 14, 20)
                        ws.set_column(15, 15, 20)
                        ws.set_column(16, 16, 20)
                        ws.set_column(17, 17, 20)
                        ws.set_column(18, 18, 20)
                        ws.set_column(19, 19, 20)
                        ws.set_column(20, 20, 20)

                        #                   ws.columm_dimensions['A'].width = 20

                        # formatotitulo = workbook.add_format(
                        #     {'bold': 1, 'text_wrap': True, 'border': 1, 'align': 'center', 'valign': 'middle',
                        #      'fg_color': '#A2D0EC'})
                        formatotitulo_filtros = workbook.add_format(
                            {'bold': 1, 'text_wrap': True, 'border': 1, 'fg_color': '#EBF5FB'})

                        formatoceldacab = workbook.add_format(
                            {'align': 'center', 'bold': 1, 'border': 1, 'text_wrap': True, 'fg_color': '#0b2f44', 'font_color':'white'})
                        formatoceldaleft = workbook.add_format(
                            {'text_wrap': True, 'align': 'center', 'valign': 'vcenter', 'border': 1})

                        ws.write(0, 0, 'N°', formatoceldacab)
                        ws.write(0, 1, 'PROGRAMA', formatoceldacab)
                        ws.write(0, 2, 'TIPO PROYECTO', formatoceldacab)
                        ws.write(0, 3, 'ÁREA DE CONOCIMIENTO', formatoceldacab)
                        ws.write(0, 4, 'SUB-ÁREA DE CONOCIMIENTO', formatoceldacab)
                        ws.write(0, 5, 'SUB-ÁREA ESPECÍFICA', formatoceldacab)
                        ws.write(0, 6, 'NOMBRE DEL PROYECTO', formatoceldacab)
                        ws.write(0, 7, 'LÍNEA DE INVESTIGACIÓN - SUB-LÍNEA ', formatoceldacab)
                        ws.write(0, 8, 'ZONA', formatoceldacab)
                        ws.write(0, 9, 'CANTÓN', formatoceldacab)
                        ws.write(0, 10, 'DISTRITO', formatoceldacab)
                        ws.write(0, 11, 'CIRCUITO', formatoceldacab)
                        ws.write(0, 12, 'SECTOR Y COORDENADAS', formatoceldacab)
                        ws.write(0, 13, 'FECHA INICIO', formatoceldacab)
                        ws.write(0, 14, 'FECHA FIN PLANIFICADA', formatoceldacab)
                        ws.write(0, 15, 'FECHA REAL', formatoceldacab)
                        ws.write(0, 16, 'HORAS DE DURACIÓN', formatoceldacab)
                        ws.write(0, 17, 'MESES DE DURACIÓN', formatoceldacab)
                        ws.write(0, 18, 'PLANIFICADO', formatoceldacab)
                        ws.write(0, 19, 'APORTES EXTERNOS', formatoceldacab)
                        ws.write(0, 20, 'TOTAL', formatoceldacab)

                        listaproyectos = ProyectosInvestigacion.objects.filter(status=True, tipo=tipoproyectos)


                        filas_recorridas = 2
                        for proyecto in listaproyectos:

                            zona = ""
                            canton = ""
                            for zon in proyecto.zona.all():
                                zona = zon.nombre
                            for canto in proyecto.canton.all():
                                canton = canto.nombre

                            ws.write('A%s' % filas_recorridas, str(proyecto.id), formatoceldaleft)
                            ws.write('B%s' % filas_recorridas, str(proyecto.programa.nombre if proyecto.programa else "NO REGISTRA"), formatoceldaleft)
                            ws.write('C%s' % filas_recorridas, str(proyecto.get_tipo_display()), formatoceldaleft)
                            ws.write('D%s' % filas_recorridas, str(proyecto.areaconocimiento.nombre if proyecto.areaconocimiento else "NO REGISTRA"), formatoceldaleft)
                            ws.write('E%s' % filas_recorridas, str(proyecto.subareaconocimiento.nombre if proyecto.subareaconocimiento else "NO REGISTRA"), formatoceldaleft)
                            ws.write('F%s' % filas_recorridas, str(proyecto.subareaespecificaconocimiento.nombre if proyecto.subareaespecificaconocimiento else "NO REGISTRA"), formatoceldaleft)
                            ws.write('G%s' % filas_recorridas, str(proyecto.nombre if proyecto.nombre else "NO REGISTRA"), formatoceldaleft)
                            ws.write('H%s' % filas_recorridas, str(proyecto.sublineainvestigacion if proyecto.sublineainvestigacion else "NO REGISTRA"), formatoceldaleft)
                            ws.write('I%s' % filas_recorridas, str(zona if proyecto.zona.exists() else "NO REGISTRA"), formatoceldaleft)
                            ws.write('J%s' % filas_recorridas, str(canton if proyecto.canton.exists() else "NO REGISTRA"), formatoceldaleft)
                            ws.write('K%s' % filas_recorridas, str(proyecto.distrito if proyecto.distrito else "NO REGISTRA"), formatoceldaleft)
                            ws.write('L%s' % filas_recorridas, str(proyecto.circuito if proyecto.circuito else "NO REGISTRA"), formatoceldaleft)
                            ws.write('M%s' % filas_recorridas, str(proyecto.sectorcoordenada if proyecto.sectorcoordenada else "NO REGISTRA"), formatoceldaleft)
                            ws.write('N%s' % filas_recorridas, str(proyecto.fechainicio if proyecto.fechainicio else "NO REGISTRA"), formatoceldaleft)
                            ws.write('O%s' % filas_recorridas, str(proyecto.fechaplaneacion if proyecto.fechaplaneacion else "NO REGISTRA"), formatoceldaleft)
                            ws.write('P%s' % filas_recorridas, str(proyecto.fechareal if proyecto.fechareal else "NO REGISTRA"), formatoceldaleft)
                            ws.write('Q%s' % filas_recorridas, str(proyecto.tiempo_duracion_horas if proyecto.tiempo_duracion_horas else "NO REGISTRA"), formatoceldaleft)
                            ws.write('R%s' % filas_recorridas, str(proyecto.tiempoejecucion if proyecto.tiempoejecucion else "NO REGISTRA"), formatoceldaleft)
                            ws.write('S%s' % filas_recorridas, str(proyecto.valorpresupuestointerno), formatoceldaleft)
                            ws.write('T%s' % filas_recorridas, 0.00, formatoceldaleft)
                            ws.write('U%s' % filas_recorridas, str(proyecto.valorpresupuestointerno), formatoceldaleft)

                            filas_recorridas += 1

                        workbook.close()
                        output.seek(0)
                        filename = 'Matriz_Proyectos_Investigacion.xlsx'
                        response = HttpResponse(output,
                                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                        response['Content-Disposition'] = 'attachment; filename=%s' % filename
                        return response
                    except Exception as ex:
                       pass

            elif action == 'excelEstudianteConvocatoria':
                try:
                    idConvocatoria = request.GET['conv']

                    __author__ = 'Unemi'
                    style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
                    style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
                    style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
                    title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                    style1 = easyxf(num_format_str='D-MMM-YY')
                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('exp_xls_post_part')
                    ws.write_merge(0, 0, 0, 10, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                    response = HttpResponse(content_type="application/ms-excel")
                    response[
                        'Content-Disposition'] = 'attachment; filename=Matriz_Participantes_Proyectos' + random.randint(
                        1, 10000).__str__() + '.xls'

                    columns = [
                        (u"NUM", 2000),
                        (u"PROYECTO", 10000),
                        (u"CEDULA", 10000),
                        (u"APELLIDOS / NOMBRES", 10000),
                        (u"CARRERA", 10000),
                        (u"NIVEL", 10000),
                        (u"CORREO PERSONAL", 10000),
                        (u"CORREO INSTITUCIONAL", 10000),
                        (u"CELULAR", 10000),
                        (u"HORAS REALIZADAS", 10000),
                    ]
                    row_num = 3
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], font_style)
                        ws.col(col_num).width = columns[col_num][1]
                    date_format = xlwt.XFStyle()
                    date_format.num_format_str = 'yyyy/mm/dd'
                    convocatoria = FechaProyectos.objects.get(id=idConvocatoria)
                    proyectos = ProyectosInvestigacion.objects.values_list('id', flat=True).filter(status=True,
                                                                                                   convocatoria=convocatoria)
                    estudiantes = ParticipantesMatrices.objects.filter(status=True, inscripcion__isnull=False,proyecto_id__in=proyectos).order_by('inscripcion__persona__apellido1', 'inscripcion__persona__apellido2')
                    row_num = 4
                    i = 0
                    for participantes in estudiantes:
                        i+=1
                        ws.write(row_num, 0, i, font_style2)
                        ws.write(row_num, 1, participantes.proyecto.nombre, font_style2)
                        ws.write(row_num, 2, participantes.inscripcion.persona.cedula, font_style2)
                        ws.write(row_num, 3, participantes.inscripcion.persona.nombre_completo_inverso(), font_style2)
                        ws.write(row_num, 4, participantes.inscripcion.carrera.nombre, font_style2)
                        ws.write(row_num, 5, str(participantes.inscripcion.nivelinscripcionmalla()), font_style2)
                        ws.write(row_num, 6, participantes.inscripcion.persona.email, font_style2)
                        ws.write(row_num, 7, participantes.inscripcion.persona.emailinst, font_style2)
                        ws.write(row_num, 8, participantes.inscripcion.persona.telefono, font_style2)
                        ws.write(row_num, 9, participantes.horas, font_style2)
                        row_num += 1
                    wb.save(response)
                    return response
                except Exception as ex:
                    pass

            # elif action == 'excelproyectosinvestigacionvinculacionD':
            #     try:
            #         tipoproyectos = request.GET['tipo']
            #         __author__ = 'Unemi'
            #         style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
            #         style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
            #         style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
            #         title = easyxf('font: name Times New Roman, color-index black, bold on , height 350; alignment: horiz centre')
            #         style1 = easyxf(num_format_str='D-MMM-YY')
            #         font_style = XFStyle()
            #         font_style.font.bold = True
            #         font_style2 = XFStyle()
            #         font_style2.font.bold = False
            #         wb = Workbook(encoding='utf-8')
            #         ws = wb.add_sheet('exp_xls_post_part')
            #         #ws.write_merge(0, 0, 0, 10, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
            #         response = HttpResponse(content_type="application/ms-excel")
            #         response[
            #             'Content-Disposition'] = 'attachment; filename=Matriz_Proyectos_Investigacion' + random.randint(
            #             1, 10000).__str__() + '.xls'
            #
            #         columns = [
            #             (u"N°", 2000),
            #             (u"PROGRAMA", 10000),
            #             (u"TIPO PROYECTO", 5000),
            #             (u"ÁREA DE CONOCIMIENTO", 18000),
            #             (u"SUB-ÁREA DE CONOCIMIENTO", 18000),
            #             (u"SUB-ÁREA ESPECÍFICA", 30000),
            #             (u"NOMBRE DEL PROYECTO", 30000),
            #             (u"LÍNEA DE INVESTIGACIÓN - SUB-LÍNEA ", 30000),
            #             (u"ZONA", 5000),
            #             (u"CANTÓN", 5000),
            #             (u"DISTRITO", 5000),
            #             (u"CIRCUITO", 5000),
            #             (u"SECTOR Y COORDENADAS", 5000),
            #             (u"FECHA INICIO", 5000),
            #             (u"FECHA FIN PLANIFICADA", 5000),
            #             (u"FECHA FIN REAL", 5000),
            #             (u"HORAS DE DURACIÓN", 5000),
            #             (u"MESES DE DURACIÓN", 5000),
            #             (u"PLANIFICADO",5000),
            #             (u"APORTES EXTERNOS", 5000),
            #             (u"TOTAL", 5000),
            #         ]
            #         row_num = 0
            #         for col_num in range(len(columns)):
            #             ws.write(row_num, col_num, columns[col_num][0], font_style)
            #             ws.col(col_num).width = columns[col_num][1]
            #         date_format = xlwt.XFStyle()
            #         date_format.num_format_str = 'yyyy/mm/dd'
            #         listaproyectos = ProyectosInvestigacion.objects.filter(status=True, tipo=tipoproyectos)
            #         mensaje='NO REGISTRA'
            #         row_num = 1
            #         for proyecto in listaproyectos:
            #             #total=float(proyecto.valorpresupuestointerno)+float(proyecto.valorpresupuestoexterno)
            #             zona=""
            #             canton=""
            #             for zon in proyecto.zona.all():
            #                 zona = zon.nombre
            #             for canto in proyecto.canton.all():
            #                 canton = canto.nombre
            #
            #             i = 0
            #             ws.write(row_num, 0, proyecto.id, font_style2)
            #             ws.write(row_num, 1, (proyecto.programa.nombre if proyecto.programa else mensaje ), font_style2)
            #             ws.write(row_num, 2, proyecto.get_tipo_display(), font_style2)
            #             ws.write(row_num, 3, (proyecto.areaconocimiento.nombre if proyecto.areaconocimiento else mensaje ), font_style2)
            #             ws.write(row_num, 4, (proyecto.subareaconocimiento.nombre if proyecto.subareaconocimiento else mensaje ), font_style2)
            #             ws.write(row_num, 5, (proyecto.subareaespecificaconocimiento.nombre if proyecto.subareaespecificaconocimiento else mensaje ), font_style2)
            #             ws.write(row_num, 6, (proyecto.nombre if proyecto.nombre else mensaje), font_style2)
            #             ws.write(row_num, 7, str(proyecto.sublineainvestigacion if proyecto.sublineainvestigacion else mensaje), font_style2)
            #             ws.write(row_num, 8,  zona if proyecto.zona.exists() else mensaje , font_style2)
            #             ws.write(row_num, 9,  canton if proyecto.canton.exists() else mensaje, font_style2)
            #             ws.write(row_num, 10, (proyecto.distrito if proyecto.distrito else mensaje), font_style2)
            #             ws.write(row_num, 11, (proyecto.circuito if proyecto.circuito else mensaje), font_style2)
            #             ws.write(row_num, 12, (proyecto.sectorcoordenada if proyecto.sectorcoordenada else mensaje ), font_style2)
            #             ws.write(row_num, 13, (proyecto.fechainicio if proyecto.fechainicio else mensaje), style1)
            #             ws.write(row_num, 14, (proyecto.fechaplaneacion if proyecto.fechaplaneacion else mensaje), style1)
            #             ws.write(row_num, 15, (proyecto.fechareal if proyecto.fechareal else mensaje), style1)
            #             ws.write(row_num, 16, (proyecto.tiempo_duracion_horas if proyecto.tiempo_duracion_horas else mensaje), font_style2)
            #             ws.write(row_num, 17, (proyecto.tiempoejecucion if proyecto.tiempoejecucion else mensaje), font_style2)
            #             ws.write(row_num, 18, proyecto.valorpresupuestointerno, font_style2)
            #             ws.write(row_num, 19,  0.00 , font_style2)
            #             ws.write(row_num, 20, proyecto.valorpresupuestointerno , font_style2)
            #             row_num += 1
            #         wb.save(response)
            #         return response
            #     except Exception as ex:
            #         pass

            elif action == 'excelinformacionproyectos':
                try:
                    tipoproyectos = request.GET['tipo']
                    output = io.BytesIO()
                    workbook = xlsxwriter.Workbook(output)
                    ws = workbook.add_worksheet('información_proyectos')
                    ws.set_column(0, 40, 30)

                    formatoceldacab = workbook.add_format(
                        {'align': 'center', 'bold': 1, 'border': 1, 'text_wrap': True, 'fg_color': '#0b2f44', 'font_color': 'white'})
                    formatoceldacab2 = workbook.add_format(
                        {'align': 'center', 'bold': 1, 'border': 1, 'text_wrap': True, 'fg_color': '#275a8a', 'font_color': 'white'})
                    formatoceldacab3 = workbook.add_format(
                        {'align': 'left', 'bold': 1, 'border': 1, 'text_wrap': True, 'fg_color': '#306da6', 'font_color': 'white'})
                    formatoceldaleft = workbook.add_format({'text_wrap': True, 'align': 'center', 'valign': 'vcenter', 'border': 1})

                    ws.merge_range('A1:U1', 'INFORMACION DE PROYECTO DE VINCULACION', formatoceldacab)
                    ws.write('V1', 'DOCENTES', formatoceldacab)
                    ws.write('W1', 'CARRERAS', formatoceldacab)
                    ws.merge_range('X1:AH1', 'BENEFICIARIOS', formatoceldacab3)
                    ws.write('A2', 'N', formatoceldacab2)
                    ws.write('B2', 'PROGRAMA', formatoceldacab2)
                    ws.write('C2', 'TIPO', formatoceldacab2)
                    ws.write('D2', 'AREA CONOCIMIENTO', formatoceldacab2)
                    ws.write('E2', 'SUB AREA CONOCIMIENTO', formatoceldacab2)
                    ws.write('F2', 'SUB AREA ESPECIFICA', formatoceldacab2)
                    ws.write('G2', 'NOMBRE PROYECTO', formatoceldacab2)
                    ws.write('H2', 'LIN - SUBLIN', formatoceldacab2)
                    ws.write('I2', 'ZONA', formatoceldacab2)
                    ws.write('J2', 'CANTON', formatoceldacab2)
                    ws.write('K2', 'DISTRITO', formatoceldacab2)
                    ws.write('L2', 'CIRCUITO', formatoceldacab2)
                    ws.write('M2', 'SECTOR-COORDENADAS', formatoceldacab2)
                    ws.write('N2', 'FECHA INICIO', formatoceldacab2)
                    ws.write('O2', 'FECHA FIN PLANIFICADA', formatoceldacab2)
                    ws.write('P2', 'FECHA FIN REAL', formatoceldacab2)
                    ws.write('Q2', 'HORAS', formatoceldacab2)
                    ws.write('R2', 'MESES', formatoceldacab2)
                    ws.write('S2', 'PLANIFICADOS', formatoceldacab2)
                    ws.write('T2', 'APORTES EX', formatoceldacab2)
                    ws.write('U2', 'TOTAL', formatoceldacab2)
                    ws.write('V2', 'NOMBRES', formatoceldacab2)
                    ws.write('W2', 'NOMBRES', formatoceldacab2)
                    ws.write('X2', 'NOMBRES', formatoceldacab3)
                    ws.write('Y2', 'CARACTERISTICA', formatoceldacab3)
                    ws.write('Z2', 'DIRECCION', formatoceldacab3)
                    ws.write('AA2', 'REPRESENTANTE', formatoceldacab3)
                    ws.write('AB2', 'CARGO', formatoceldacab3)
                    ws.write('AC2', 'TLFN', formatoceldacab3)
                    ws.write('AD2', 'CORREO', formatoceldacab3)
                    ws.write('AE2', 'COORDENADAS', formatoceldacab3)
                    ws.write('AF2', '# BENEF DIRECTO', formatoceldacab3)
                    ws.write('AG2', '# BENEF INDIRECTO', formatoceldacab3)
                    ws.write('AH2', 'ARCHIVO', formatoceldacab3)

                    listaproyectos = ProyectosInvestigacion.objects.filter(status=True, tipo=tipoproyectos)
                    mensaje = 'NO REGISTRA'
                    fila_docente = 2
                    fila_carrera = 2
                    fila_beneficiario = 2
                    fila = 2
                    for proyecto in listaproyectos:
                        docentes = ParticipantesMatrices.objects.filter(status=True, tipoparticipante__tipo=1, proyecto_id=proyecto.id).order_by('fecha_creacion')
                        carreras = CarrerasParticipantes.objects.filter(status=True, proyecto_id=proyecto.id)
                        beneficiarios = Beneficiarios.objects.filter(status=True, proyecto_id=proyecto.id)
                        zona = ""
                        canton = ""
                        for zon in proyecto.zona.all():
                            zona = zon.nombre
                        for canto in proyecto.canton.all():
                            canton = canto.nombre
                        ws.write(fila, 0, proyecto.id, formatoceldaleft)
                        ws.write(fila, 1, (proyecto.programa.nombre if proyecto.programa else mensaje), formatoceldaleft)
                        ws.write(fila, 2, (proyecto.get_tipo_display()), formatoceldaleft)
                        ws.write(fila, 3, (proyecto.areaconocimiento.nombre if proyecto.areaconocimiento else mensaje), formatoceldaleft)
                        ws.write(fila, 4, (proyecto.subareaconocimiento.nombre if proyecto.subareaconocimiento else mensaje), formatoceldaleft)
                        ws.write(fila, 5, (proyecto.subareaespecificaconocimiento.nombre if proyecto.subareaespecificaconocimiento else mensaje), formatoceldaleft)
                        ws.write(fila, 6, (proyecto.nombre if proyecto.nombre else mensaje), formatoceldaleft)
                        ws.write(fila, 7, str(proyecto.sublineainvestigacion if proyecto.sublineainvestigacion else mensaje), formatoceldaleft)
                        ws.write(fila, 8, zona if proyecto.zona.exists() else mensaje, formatoceldaleft)
                        ws.write(fila, 9, canton if proyecto.canton.exists() else mensaje, formatoceldaleft)
                        ws.write(fila, 10, (proyecto.distrito if proyecto.distrito else mensaje), formatoceldaleft)
                        ws.write(fila, 11, (proyecto.circuito if proyecto.circuito else mensaje), formatoceldaleft)
                        ws.write(fila, 12, (proyecto.sectorcoordenada if proyecto.sectorcoordenada else mensaje), formatoceldaleft)
                        ws.write(fila, 13, (proyecto.fechainicio if proyecto.fechainicio else mensaje), formatoceldaleft)
                        ws.write(fila, 14, (proyecto.fechaplaneacion if proyecto.fechaplaneacion else mensaje), formatoceldaleft)
                        ws.write(fila, 15, (proyecto.fechareal if proyecto.fechareal else mensaje), formatoceldaleft)
                        ws.write(fila, 16, (proyecto.tiempo_duracion_horas if proyecto.tiempo_duracion_horas else mensaje), formatoceldaleft)
                        ws.write(fila, 17, (proyecto.tiempoejecucion if proyecto.tiempoejecucion else mensaje), formatoceldaleft)
                        ws.write(fila, 18, proyecto.valorpresupuestointerno, formatoceldaleft)
                        ws.write(fila, 19, 0.00, formatoceldaleft)
                        ws.write(fila, 20, proyecto.valorpresupuestointerno, formatoceldaleft)

                        for docente in docentes:
                            if docente:
                                ws.write(fila_docente, 21, str(docente.profesor), formatoceldaleft)
                                fila_docente+=1

                        for carrera in carreras:
                            if carrera:
                                ws.write(fila_carrera, 22, (str(carrera.carrera.nombre) if carrera.carrera else mensaje), formatoceldaleft)
                                fila_carrera+=1

                        for beneficiario in beneficiarios:
                            if beneficiario:
                                ws.write(fila_beneficiario, 23, (str(beneficiario.nombre) if beneficiario.nombre else mensaje), formatoceldaleft)
                                ws.write(fila_beneficiario, 24, (str(beneficiario.caracteristica) if beneficiario.caracteristica else mensaje), formatoceldaleft)
                                ws.write(fila_beneficiario, 25, (str(beneficiario.direccion) if beneficiario.direccion else mensaje), formatoceldaleft)
                                ws.write(fila_beneficiario, 26, (str(beneficiario.representante) if beneficiario.representante else mensaje), formatoceldaleft)
                                ws.write(fila_beneficiario, 27, (str(beneficiario.cargo_repre.descripcion) if beneficiario.cargo_repre.descripcion else mensaje), formatoceldaleft)
                                ws.write(fila_beneficiario, 28, (str(beneficiario.telefono) if beneficiario.telefono else mensaje), formatoceldaleft)
                                ws.write(fila_beneficiario, 29, (str(beneficiario.correo) if beneficiario.correo else mensaje), formatoceldaleft)
                                ws.write(fila_beneficiario, 30, (str(beneficiario.coordenadas) if beneficiario.coordenadas else mensaje), formatoceldaleft)
                                ws.write(fila_beneficiario, 31, (str(beneficiario.num_beneficiario_directo) if beneficiario.num_beneficiario_directo else mensaje), formatoceldaleft)
                                ws.write(fila_beneficiario, 32, (str(beneficiario.num_beneficiario_indirecto) if beneficiario.num_beneficiario_indirecto else mensaje), formatoceldaleft)
                                ws.write(fila_beneficiario, 33, (str(beneficiario.archivo) if beneficiario.archivo else mensaje), formatoceldaleft)
                                fila_beneficiario+=1

                        if fila_docente >= fila_carrera:
                            fila = fila_docente
                            fila_beneficiario = fila_docente
                        elif fila_carrera >= fila_docente:
                            fila = fila_carrera
                            fila_beneficiario = fila_carrera
                        else:
                            fila+=1


                    # filasdocenterecorrida = 0
                    # for docente in listadocentes.filter(proyecto_id=369):
                    #     zona = ""
                    #     canton = ""
                    #     for zon in docente.proyecto.zona.all():
                    #         zona = zon.nombre
                    #     for canto in docente.proyecto.canton.all():
                    #         canton = canto.nombre
                    #
                    #     ws.write('V%s' % fila_docente, str(docente.profesor), formatoceldaleft)
                    #     fila_docente += 1
                    #     filasdocenterecorrida += 1
                    # if filasdocenterecorrida > 0:
                    #     fila_hasta = fila_desde + filasdocenterecorrida - 1
                    #     ws.merge_range('A' + str(fila_desde) + ':' + 'A' + str(fila_hasta), docente.proyecto.id, formatoceldaleft)
                    #     ws.merge_range('B' + str(fila_desde) + ':' + 'B' + str(fila_hasta), docente.proyecto.programa.nombre if docente.proyecto.programa.nombre else mensaje, formatoceldaleft)
                    #     ws.merge_range('C' + str(fila_desde) + ':' + 'C' + str(fila_hasta), docente.proyecto.get_tipo_display() if docente.proyecto.get_tipo_display() else mensaje, formatoceldaleft)
                    #     ws.merge_range('D' + str(fila_desde) + ':' + 'D' + str(fila_hasta), docente.proyecto.areaconocimiento.nombre if docente.proyecto.areaconocimiento.nombre else mensaje, formatoceldaleft)
                    #     ws.merge_range('E' + str(fila_desde) + ':' + 'E' + str(fila_hasta), docente.proyecto.subareaconocimiento.nombre if docente.proyecto.subareaconocimiento.nombre else mensaje, formatoceldaleft)
                    #     ws.merge_range('F' + str(fila_desde) + ':' + 'F' + str(fila_hasta), docente.proyecto.subareaespecificaconocimiento.nombre if docente.proyecto.subareaespecificaconocimiento.nombre else mensaje, formatoceldaleft)
                    #     ws.merge_range('G' + str(fila_desde) + ':' + 'G' + str(fila_hasta), docente.proyecto.nombre if docente.proyecto.nombre else mensaje, formatoceldaleft)
                    #     ws.merge_range('H' + str(fila_desde) + ':' + 'H' + str(fila_hasta), docente.proyecto.sublineainvestigacion.nombre if docente.proyecto.sublineainvestigacion.nombre else mensaje, formatoceldaleft)
                    #     ws.merge_range('I' + str(fila_desde) + ':' + 'I' + str(fila_hasta), zona if docente.proyecto.zona.exists() else mensaje, formatoceldaleft)
                    #     ws.merge_range('J' + str(fila_desde) + ':' + 'J' + str(fila_hasta), canton if docente.proyecto.canton.exists() else mensaje, formatoceldaleft)
                    #     ws.merge_range('K' + str(fila_desde) + ':' + 'K' + str(fila_hasta), docente.proyecto.distrito if docente.proyecto.distrito else mensaje, formatoceldaleft)
                    #     ws.merge_range('L' + str(fila_desde) + ':' + 'L' + str(fila_hasta), docente.proyecto.circuito if docente.proyecto.circuito else mensaje, formatoceldaleft)
                    #     ws.merge_range('M' + str(fila_desde) + ':' + 'M' + str(fila_hasta), docente.proyecto.sectorcoordenada if docente.proyecto.sectorcoordenada else mensaje, formatoceldaleft)
                    #     ws.merge_range('N' + str(fila_desde) + ':' + 'N' + str(fila_hasta), docente.proyecto.fechainicio if docente.proyecto.fechainicio else mensaje, formatoceldaleft)
                    #     ws.merge_range('O' + str(fila_desde) + ':' + 'O' + str(fila_hasta), docente.proyecto.fechaplaneacion if docente.proyecto.fechaplaneacion else mensaje, formatoceldaleft)
                    #     ws.merge_range('P' + str(fila_desde) + ':' + 'P' + str(fila_hasta), docente.proyecto.fechareal if docente.proyecto.fechareal else mensaje, formatoceldaleft)
                    #     ws.merge_range('Q' + str(fila_desde) + ':' + 'Q' + str(fila_hasta), docente.proyecto.tiempo_duracion_horas if docente.proyecto.tiempo_duracion_horas else mensaje, formatoceldaleft)
                    #     fila_desde = fila_hasta + 1
                    # fila_docente += 1
                    workbook.close()
                    output.seek(0)
                    filename = 'Información_de_proyectos_de_vinculación.xlsx'
                    response = HttpResponse(output,
                                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    response['Content-Disposition'] = 'attachment; filename=%s' % filename
                    return response
                except Exception as ex:
                    print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
                    pass

            elif action == 'excelpresupuesto':
                    try:
                        __author__ = 'Unemi'

                        proyecto = ProyectosInvestigacion.objects.get(pk=request.GET['pry'])

                        output = io.BytesIO()
                        workbook = xlsxwriter.Workbook(output)
                        ws = workbook.add_worksheet('Por actividad')
                        ws.set_column(0, 0, 10)
                        ws.set_column(1, 1, 40)
                        ws.set_column(2, 2, 40)
                        ws.set_column(3, 3, 10)
                        ws.set_column(4, 4, 10)
                        ws.set_column(5, 5, 10)
                        ws.set_column(6, 6, 10)
                        ws.set_column(7, 7, 10)
                        ws.set_column(8, 8, 10)
                        ws.set_column(9, 9, 10)



                        formatotitulo_filtros = workbook.add_format({'bold': 1, 'text_wrap': True, 'border': 1, 'fg_color': '#EBF5FB'})

                        formatoceldacab = workbook.add_format({'align': 'center', 'bold': 1, 'border': 1, 'text_wrap': True, 'fg_color': '#0b2f44', 'font_color':'white'})
                        formatoceldaleft = workbook.add_format({'text_wrap': True, 'align': 'center', 'valign': 'vcenter', 'border': 1})

                        ws.write(0, 0, 'Cant.', formatoceldacab)
                        ws.write(0, 1, 'Rubro', formatoceldacab)
                        ws.write(0, 2, 'Especificaciones', formatoceldacab)
                        ws.write(0, 3, 'C. Unitario', formatoceldacab)
                        ws.write(0, 4, 'Subtotal', formatoceldacab)
                        ws.write(0, 5, 'Iva', formatoceldacab)
                        ws.write(0, 6, 'Total', formatoceldacab)

                        acciones = MarcoLogico.objects.filter(status=True, proyecto=proyecto, arbolObjetivo__tipo=2, arbolObjetivo__parentID__isnull=False).order_by('arbolObjetivo__orden')
                        filas_recorridas = 2

                        for accion in acciones:
                            i = 0
                            ws.merge_range('A%s:G%s' %(filas_recorridas,filas_recorridas), str(accion.resumen_narrativo), formatoceldaleft)
                            filas_recorridas += 1
                            for rubro in accion.suministro():
                                ws.write('A%s' % filas_recorridas, (rubro.cantidad), formatoceldaleft)
                                ws.write('B%s' % filas_recorridas, str(rubro.suministro.rubro if rubro.suministro else rubro.producto), formatoceldaleft)
                                ws.write('C%s' % filas_recorridas, str(rubro.especificaciones), formatoceldaleft)
                                ws.write('D%s' % filas_recorridas, (null_to_decimal(rubro.costo_unitario,4)), formatoceldaleft)
                                ws.write_formula('E%s'% filas_recorridas, '=(A%s*D%s)' % (filas_recorridas,filas_recorridas), formatoceldaleft)
                                ws.write_formula('F%s'% filas_recorridas, '=(E%s*0.12)' % (filas_recorridas) if rubro.suministro.aplicaIva else '0', formatoceldaleft)
                                ws.write_formula('G%s'% filas_recorridas, '=SUM(E%s+F%s)' % (filas_recorridas,filas_recorridas), formatoceldaleft)

                                filas_recorridas += 1

                        ws.write('F%s'% filas_recorridas, 'Total',formatoceldaleft)
                        ws.write_formula('G%s'% filas_recorridas, '=SUM(G2:G%s)' % (filas_recorridas-1), formatoceldaleft)

                        ws2 = workbook.add_worksheet('Consolidado')

                        ws2.set_column(0, 0, 10)
                        ws2.set_column(1, 1, 40)
                        ws2.set_column(2, 2, 40)
                        ws2.set_column(3, 3, 10)
                        ws2.set_column(4, 4, 10)
                        ws2.set_column(5, 5, 10)
                        ws2.set_column(6, 6, 10)

                        ws2.write(0, 0, 'Cant.', formatoceldacab)
                        ws2.write(0, 1, 'Rubro', formatoceldacab)
                        ws2.write(0, 2, 'Especificaciones', formatoceldacab)
                        ws2.write(0, 3, 'C. Unitario', formatoceldacab)
                        ws2.write(0, 4, 'Subtotal', formatoceldacab)
                        ws2.write(0, 5, 'Iva', formatoceldacab)
                        ws2.write(0, 6, 'Total', formatoceldacab)

                        presupuesto = Presupuesto.objects.filter(status=True, proyecto=proyecto).distinct('suministro__pk')
                        filas_recorridas = 2
                        for pre in presupuesto:
                            i = 0
                            presupuesto = Presupuesto.objects.filter(status=True, proyecto=proyecto,suministro=pre.suministro)
                            ws2.write('A%s' % filas_recorridas, presupuesto.aggregate(valor=Sum('cantidad'))['valor'], formatoceldaleft)
                            ws2.write('B%s' % filas_recorridas,str(pre.suministro.rubro if pre.suministro else pre.producto),formatoceldaleft)
                            ws2.write('C%s' % filas_recorridas, str(pre.especificaciones), formatoceldaleft)
                            ws2.write('D%s' % filas_recorridas, (null_to_decimal(pre.costo_unitario, 4)),formatoceldaleft)
                            ws2.write_formula('E%s' % filas_recorridas,'=(A%s*D%s)' % (filas_recorridas, filas_recorridas), formatoceldaleft)
                            ws2.write_formula('F%s' % filas_recorridas,'=(E%s*0.12)' % (filas_recorridas) if pre.suministro.aplicaIva else '0',formatoceldaleft)
                            ws2.write_formula('G%s' % filas_recorridas,'=SUM(E%s+F%s)' % (filas_recorridas, filas_recorridas), formatoceldaleft)

                            filas_recorridas += 1
                        ws2.write('F%s' % filas_recorridas, 'Total', formatoceldaleft)
                        ws2.write_formula('G%s' % filas_recorridas, '=SUM(G2:G%s)' % (filas_recorridas - 1),formatoceldaleft)
                        workbook.close()
                        output.seek(0)
                        filename = 'Presupuesto.xlsx'
                        response = HttpResponse(output,content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                        response['Content-Disposition'] = 'attachment; filename=%s' % filename
                        return response
                    except Exception as ex:
                       pass

            elif action == 'excelpresupuestoconvocatoria':
                    try:
                        __author__ = 'Unemi'

                        if request.GET['est'] == 0 :
                            proyecto = ProyectosInvestigacion.objects.values_list('id',flat=True).filter(convocatoria = request.GET['id'], status= True)
                        else:
                            proyecto = ProyectosInvestigacion.objects.values_list('id',flat=True).filter(convocatoria=request.GET['id'],aprobacion=request.GET['est'], status= True)

                        output = io.BytesIO()
                        workbook = xlsxwriter.Workbook(output)

                        formatotitulo_filtros = workbook.add_format({'bold': 1, 'text_wrap': True, 'border': 1, 'fg_color': '#EBF5FB'})

                        formatoceldacab = workbook.add_format({'align': 'center', 'bold': 1, 'border': 1, 'text_wrap': True, 'fg_color': '#0b2f44', 'font_color':'white'})
                        formatoceldaleft = workbook.add_format({'text_wrap': True, 'align': 'center', 'valign': 'vcenter', 'border': 1})

                        acciones = MarcoLogico.objects.filter(status=True, proyecto=proyecto, arbolObjetivo__tipo=2, arbolObjetivo__parentID__isnull=False).order_by('arbolObjetivo__orden')
                        filas_recorridas = 2

                        ws2 = workbook.add_worksheet('Consolidado')

                        ws2.set_column(0, 0, 10)
                        ws2.set_column(1, 1, 40)
                        ws2.set_column(2, 2, 40)
                        ws2.set_column(3, 3, 10)
                        ws2.set_column(4, 4, 10)
                        ws2.set_column(5, 5, 10)
                        ws2.set_column(6, 6, 10)

                        ws2.write(0, 0, 'Cant.', formatoceldacab)
                        ws2.write(0, 1, 'Rubro', formatoceldacab)
                        ws2.write(0, 2, 'Especificaciones', formatoceldacab)
                        ws2.write(0, 3, 'C. Unitario', formatoceldacab)
                        ws2.write(0, 4, 'Subtotal', formatoceldacab)
                        ws2.write(0, 5, 'Iva', formatoceldacab)
                        ws2.write(0, 6, 'Total', formatoceldacab)

                        presupuesto = Presupuesto.objects.filter(status=True, proyecto__id__in=proyecto).distinct('suministro__pk')
                        filas_recorridas = 2
                        for pre in presupuesto:
                            i = 0
                            presupuesto = Presupuesto.objects.filter(status=True, proyecto__id__in=proyecto,suministro=pre.suministro)
                            ws2.write('A%s' % filas_recorridas, presupuesto.aggregate(valor=Sum('cantidad'))['valor'], formatoceldaleft)
                            ws2.write('B%s' % filas_recorridas,str(pre.suministro.rubro if pre.suministro else pre.producto),formatoceldaleft)
                            ws2.write('C%s' % filas_recorridas, str(pre.especificaciones), formatoceldaleft)
                            ws2.write('D%s' % filas_recorridas, (null_to_decimal(pre.costo_unitario, 4)),formatoceldaleft)
                            ws2.write_formula('E%s' % filas_recorridas,'=(A%s*D%s)' % (filas_recorridas, filas_recorridas), formatoceldaleft)
                            ws2.write_formula('F%s' % filas_recorridas,'=(E%s*0.12)' % (filas_recorridas) if pre.suministro.aplicaIva else '0',formatoceldaleft)
                            ws2.write_formula('G%s' % filas_recorridas,'=SUM(E%s+F%s)' % (filas_recorridas, filas_recorridas), formatoceldaleft)

                            filas_recorridas += 1
                        ws2.write('F%s' % filas_recorridas, 'Total', formatoceldaleft)
                        ws2.write_formula('G%s' % filas_recorridas, '=SUM(G2:G%s)' % (filas_recorridas - 1),formatoceldaleft)
                        workbook.close()
                        output.seek(0)
                        filename = 'Presupuesto.xlsx'
                        response = HttpResponse(output,content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                        response['Content-Disposition'] = 'attachment; filename=%s' % filename
                        return response
                    except Exception as ex:
                       pass

            elif action == 'excelcronograma':
                    try:
                        __author__ = 'Unemi'

                        proyecto = ProyectosInvestigacion.objects.get(pk=request.GET['pry'])

                        output = io.BytesIO()
                        workbook = xlsxwriter.Workbook(output)
                        ws = workbook.add_worksheet('cronograma')
                        ws.set_column(0, 0, 60)
                        ws.set_column(1, 1, 40)
                        ws.set_column(2, 2, 10)
                        ws.set_column(3, 3, 10)

                        #                   ws.columm_dimensions['A'].width = 20

                        # formatotitulo = workbook.add_format(
                        #     {'bold': 1, 'text_wrap': True, 'border': 1, 'align': 'center', 'valign': 'middle',
                        #      'fg_color': '#A2D0EC'})
                        formatotitulo_filtros = workbook.add_format({'bold': 1, 'text_wrap': True, 'border': 1, 'fg_color': '#EBF5FB'})
                        formatoceldacab = workbook.add_format({'align': 'center', 'bold': 1, 'border': 1, 'text_wrap': True, 'fg_color': '#0b2f44', 'font_color':'white'})
                        formatoceldaleft = workbook.add_format({'text_wrap': True, 'align': 'center', 'valign': 'vcenter', 'border': 1})

                        ws.write(0, 0, 'Tarea', formatoceldacab)
                        ws.write(0, 1, 'Responsable', formatoceldacab)
                        ws.write(0, 2, 'Inicio', formatoceldacab)
                        ws.write(0, 3, 'Fin', formatoceldacab)

                        acciones = MarcoLogico.objects.filter(status=True, proyecto=proyecto, arbolObjetivo__tipo=2, arbolObjetivo__parentID__isnull=False).order_by('arbolObjetivo__orden')
                        conograma = Cronograma.objects.filter(status=True, proyecto=proyecto)
                        filas_recorridas = 2

                        for accion in acciones:
                            i = 0
                            ws.merge_range('A%s:D%s' %(filas_recorridas,filas_recorridas), str(accion.resumen_narrativo), formatoceldacab)
                            filas_recorridas += 1
                            for crono in conograma.filter(aobjetivo=accion.arbolObjetivo) :
                                if crono.responsable.all().count()>1:
                                #ws.write('A%s' % filas_recorridas, crono.descripcion, formatoceldaleft)
                                    ws.merge_range('A%s:A%s' % (filas_recorridas, (filas_recorridas + crono.responsable.all().count()-1)),crono.descripcion, formatoceldaleft)
                                    ws.merge_range('C%s:C%s' % (filas_recorridas, (filas_recorridas + crono.responsable.all().count()-1)),str(crono.fecha_inicio), formatoceldaleft)
                                    ws.merge_range('D%s:D%s' % (filas_recorridas, (filas_recorridas + crono.responsable.all().count()-1)),str(crono.fecha_fin), formatoceldaleft)
                                else:
                                    ws.write('A%s' % filas_recorridas, crono.descripcion, formatoceldaleft)
                                    ws.write('C%s' % filas_recorridas, str(crono.fecha_inicio), formatoceldaleft)
                                    ws.write('D%s' % filas_recorridas, str(crono.fecha_fin), formatoceldaleft)
                                for res in crono.responsable.all():
                                    ws.write('B%s' % filas_recorridas, str(res.persona.nombre_completo_inverso()), formatoceldaleft)
                                    filas_recorridas += 1

                        workbook.close()
                        output.seek(0)
                        filename = 'Cronograma.xlsx'
                        response = HttpResponse(output,content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                        response['Content-Disposition'] = 'attachment; filename=%s' % filename
                        return response
                    except Exception as ex:
                       pass

            elif action == 'excelsuministro':
                    try:
                        __author__ = 'Unemi'


                        output = io.BytesIO()
                        workbook = xlsxwriter.Workbook(output)
                        ws = workbook.add_worksheet('Suministros')
                        ws.set_column(0, 0, 40)
                        ws.set_column(1, 1, 40)
                        ws.set_column(2, 2, 10)
                        ws.set_column(3, 3, 10)
                        ws.set_column(4, 4, 10)

                        formatotitulo_filtros = workbook.add_format({'bold': 1, 'text_wrap': True, 'border': 1, 'fg_color': '#EBF5FB'})

                        formatoceldacab = workbook.add_format({'align': 'center', 'bold': 1, 'border': 1, 'text_wrap': True, 'fg_color': '#0b2f44', 'font_color':'white'})
                        formatoceldaleft = workbook.add_format({'text_wrap': True, 'align': 'center', 'valign': 'vcenter', 'border': 1})

                        ws.write(0, 0, 'Rubro', formatoceldacab)
                        ws.write(0, 1, 'Especificaciones', formatoceldacab)
                        ws.write(0, 2, 'Costo', formatoceldacab)
                        ws.write(0, 3, 'Aplica IVA', formatoceldacab)
                        ws.write(0, 4, 'Estado', formatoceldacab)

                        suministros = Suministro.objects.filter(status=True)
                        filas_recorridas = 2

                        for rubro in suministros:
                            i = 0
                            ws.write('A%s' % filas_recorridas, rubro.rubro, formatoceldaleft)
                            ws.write('B%s' % filas_recorridas, rubro.especificacion, formatoceldaleft)
                            ws.write('C%s' % filas_recorridas, rubro.costo_unitario, formatoceldaleft)
                            ws.write('D%s' % filas_recorridas, 'SI' if rubro.aplicaIva else 'NO' , formatoceldaleft)
                            ws.write('E%s' % filas_recorridas, 'ACTIVO' if rubro.activo else 'INACTIVO' , formatoceldaleft)

                            filas_recorridas += 1

                        workbook.close()
                        output.seek(0)
                        filename = 'Insumos.xlsx'
                        response = HttpResponse(output,content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                        response['Content-Disposition'] = 'attachment; filename=%s' % filename
                        return response
                    except Exception as ex:
                       pass

            elif action == 'matrizparticipantes':
                try:
                    tipoproyectos = request.GET['tipo']
                    output = io.BytesIO()
                    workbook = xlsxwriter.Workbook(output)
                    ws = workbook.add_worksheet('información_proyectos')
                    ws.set_column(0, 28, 25)

                    formato_titulo = workbook.add_format(
                        {'align': 'center', 'bold': 1, 'border': 1, 'text_wrap': True, 'fg_color': '#0b2f44', 'font_color': 'white', 'size': 16})
                    formato_titulo_1 = workbook.add_format(
                        {'align': 'center', 'bold': 1, 'border': 1, 'text_wrap': True, 'fg_color': '#275a8a', 'font_color': 'white'})
                    formato_titulo_2 = workbook.add_format(
                        {'align': 'center', 'bold': 1, 'border': 1, 'text_wrap': True, 'fg_color': '#483D8B', 'font_color': 'white'})
                    formato_cabecera = workbook.add_format(
                        {'align': 'center', 'bold': 1, 'border': 1, 'text_wrap': True, 'fg_color': '#0b2f44', 'font_color': 'white'})
                    formatoceldacab3 = workbook.add_format(
                        {'align': 'left', 'bold': 1, 'border': 1, 'text_wrap': True, 'fg_color': '#306da6', 'font_color': 'white'})
                    formatoceldaleft = workbook.add_format({'text_wrap': True, 'align': 'center', 'valign': 'vcenter', 'border': 1})

                    ws.merge_range('A1:AC1', 'MATRIZ  DATOS ESTUDIANTES PRACTICANTES DE LA UNIVERSIDAD ESTATAL DE MILAGRO', formato_titulo)
                    ws.merge_range('A3:U3', 'DATOS DEL ESTUDIANTE', formato_titulo_1)
                    ws.merge_range('V3:AC3', 'DATOS DE LA INSTITUCION', formato_titulo_2)
                    ws.write('A4', 'Nº', formato_cabecera)
                    ws.write('B4', 'ZONA', formato_cabecera)
                    ws.write('C4', 'DISTRITO', formato_cabecera)
                    ws.write('D4', 'PROVINCIA', formato_cabecera)
                    ws.write('E4', 'INSTITUCIÓN DE EDUCACIÓN SUPERIOR QUE PROVIENE EL ESTUDIANTE', formato_cabecera)
                    ws.write('F4', 'NOMBRES', formato_cabecera)
                    ws.write('G4', 'CÉDULA', formato_cabecera)
                    ws.write('H4', 'CELULAR', formato_cabecera)
                    ws.write('I4', 'CORREO', formato_cabecera)
                    ws.write('J4', 'DIRECCIÓN', formato_cabecera)
                    ws.write('K4', 'CANTÓN', formato_cabecera)
                    ws.write('L4', 'CARRERA ', formato_cabecera)
                    ws.write('M4', 'SEMESTRE ', formato_cabecera)
                    ws.write('N4', 'NIVEL EDUCATIVO EN EL QUE REALIZA LAS PRÁCTICAS ', formato_cabecera)
                    ws.write('O4', '# DE HORAS DIARIAS ', formato_cabecera)
                    ws.write('P4', 'TOTAL DE HORAS', formato_cabecera)
                    ws.write('Q4', '# DE DÍAS A LA SEMANA', formato_cabecera)
                    ws.write('R4', 'TIPO DE PRÁCTICA', formato_cabecera)
                    ws.write('S4', 'PERÍODO', formato_cabecera)
                    ws.write('T4', 'NOMBRE TUTOR', formato_cabecera)
                    ws.write('U4', 'CORREO TUTOR', formato_cabecera)
                    ws.write('V4', 'NOMBRE', formato_cabecera)
                    ws.write('W4', 'AMIE', formato_cabecera)
                    ws.write('X4', 'CANTÓN', formato_cabecera)
                    ws.write('Y4', 'PARROQUIA', formato_cabecera)
                    ws.write('Z4', 'DIRECCIÓN', formato_cabecera)
                    ws.write('AA4', 'NOMBRE RECTOR', formato_cabecera)
                    ws.write('AB4', 'CELULAR RECTOR', formato_cabecera)
                    ws.write('AC4', 'CORREO INSTITUCIONAL', formato_cabecera)

                    workbook.close()
                    output.seek(0)
                    filename = 'matriz_datos_estudiantes_participantes.xlsx'
                    response = HttpResponse(output,
                                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    response['Content-Disposition'] = 'attachment; filename=%s' % filename
                    return response
                except Exception as ex:
                    pass

            elif action == 'add':
                try:
                    data['title'] = u'Adicionar Programa'
                    form = ProgramasVinculacionForm(initial={"fechainicio": datetime.now().date()})
                    data['form'] = form
                    return render(request, "inv_vinculacion/add.html", data)
                except Exception as ex:
                    pass

            if action == 'addproyecto':
                try:
                    data['title'] = u'Adicionar Proyecto'
                    form = ProyectoVinculacion1Form(initial={
                        'sectorcoordenada': "-2.1498491508722646,-79.60318654775621",
                    })
                    data['form'] = form
                    return render(request, "inv_vinculacion/addproyecto.html", data)
                except Exception as ex:
                    pass

            elif action == 'crearproyecto':
                try:
                    data['title'] = u'Formulación de proyectos de servicios comunitarios'
                    data['conv'] = request.GET['conv']
                    return render(request, "inv_vinculacion/crearproyecto.html", data)
                except Exception as ex:
                    pass

            elif action == 'addfechafinp':
                try:
                    data['title'] = u'Adicionar fecha fin del proyecto'
                    form = FechaFinProyectoFrom()
                    data['form'] = form
                    data['proyectos'] = proyectos = ProyectosInvestigacion.objects.get(pk=request.GET['id'])
                    return render(request, "inv_vinculacion/add_fechafin.html", data)
                except Exception as ex:
                    pass

            elif action == 'archivoenviados':
                try:
                    data['title'] = u'Archivos Enviados'
                    if 'id' in request.GET:
                        ids = request.GET['id']
                        data['proyectoss']=ProyectosInvestigacion.objects.get(pk=ids)
                        data['aprobacion']=ProyectosInvestigacionAprobacion.objects.filter(proyecto__id=ids)
                    return render(request, "inv_vinculacion/evidenciadearchivoenviados.html", data)
                except Exception as ex:
                    pass

            elif action == 'cargarsublineas':
                try:
                    # puede_realizar_accion(request, 'sga.puede_modificar_profesor_materia')
                    if 'linea_id' in request.GET:
                        lista = []
                        lineas= SubLineaInvestigacion.objects.filter(lineainvestigacion=int(request.GET['linea_id']))
                        for lis in lineas:
                            lista.append([lis.id, lis.nombre])
                        data = {"results": "ok", 'lista':lista}
                        return JsonResponse(data)
                except Exception as ex:
                    pass

            elif action == 'cargarsublineas1':
                try:
                    # puede_realizar_accion(request, 'sga.puede_modificar_profesor_materia')
                    sublineas = LineaProgramasInvestigacion.objects.values_list('sublineainvestigacion__id',flat=True).filter(status=True,programasinvestigacion__id=int(request.GET['idp'])).order_by('-id')
                    if 'linea_id' in request.GET:
                        lista = []
                        lineas= SubLineaInvestigacion.objects.filter(lineainvestigacion=int(request.GET['linea_id']), id__in=sublineas)
                        for lis in lineas:
                            lista.append([lis.id, lis.nombre])
                        data = {"results": "ok", 'lista':lista}
                        return JsonResponse(data)
                except Exception as ex:
                    pass

            elif action == 'cargaprograma':
                try:
                    # puede_realizar_accion(request, 'sga.puede_modificar_profesor_materia')
                    lineas1 = LineaProgramasInvestigacion.objects.values_list('lineainvestigacion__id',flat=True).filter(status=True, programasinvestigacion__id=int(request.GET['idp'])).order_by('-id').distinct()
                    areas1 = AreaProgramasInvestigacion.objects.values_list('areaconocimiento__id', flat=True).filter(status=True, programasinvestigacion__id=int(request.GET['idp'])).order_by('-id').distinct()
                    lista = []
                    lista1 = []
                    lineas= LineaInvestigacion.objects.filter(id__in=lineas1)
                    for lis in lineas:
                        lista.append([lis.id, lis.nombre])

                    areas=AreaConocimientoTitulacion.objects.filter(id__in=areas1)
                    for lis in areas:
                        lista1.append([lis.id, lis.nombre])
                    data = {"results": "ok", 'lista':lista, 'lista1':lista1}
                    return JsonResponse(data)
                except Exception as ex:
                    pass
            elif action == 'cargarsublineas1':
                try:
                    # puede_realizar_accion(request, 'sga.puede_modificar_profesor_materia')
                    sublineas = LineaProgramasInvestigacion.objects.values_list('sublineainvestigacion__id',flat=True).filter(status=True,programasinvestigacion__id=int(request.GET['idp'])).order_by('-id')
                    if 'linea_id' in request.GET:
                        lista = []
                        lineas = SubLineaInvestigacion.objects.filter(lineainvestigacion=int(request.GET['linea_id']),id__in=sublineas)
                        for lis in lineas:
                            lista.append([lis.id, lis.nombre])
                        data = {"results": "ok", 'lista': lista}
                        return JsonResponse(data)
                except Exception as ex:
                    pass
            elif action == 'subareaconocimiento1':
                try:
                    # puede_realizar_accion(request, 'sga.puede_modificar_profesor_materia')

                    if 'id' in request.GET:
                        lista = []
                        subareaconocimiento = SubAreaConocimientoTitulacion.objects.filter(
                            areaconocimiento=int(request.GET['id']))
                        lista1 = []
                        subespeciconoocimiento = SubAreaEspecificaConocimientoTitulacion.objects.filter(
                            areaconocimiento=int(request.GET['id']))

                        for lis in subespeciconoocimiento:
                            lista1.append([lis.id, lis.nombre])

                        for lis in subareaconocimiento:
                            lista.append([lis.id, lis.nombre])
                        data = {"results": "ok", 'lista': lista, 'lista1': lista1}
                        return JsonResponse(data)
                except Exception as ex:
                    pass
            elif action == 'cargacanton':
                try:
                    # puede_realizar_accion(request, 'sga.puede_modificar_profesor_materia')
                    if 'zona_id' in request.GET:
                        lista = []
                        lineas= Canton.objects.filter(zona__id=int(request.GET['zona_id']))
                        for lis in lineas:
                            lista.append([lis.id, lis.nombre])
                        data = {"results": "ok", 'lista':lista}
                        return JsonResponse(data)
                except Exception as ex:
                    pass


            elif action == 'editprograma':
                try:
                    data['title'] = u'Editar Programa'
                    data['programas'] = programas = ProgramasInvestigacion.objects.get(pk=request.GET['id'])
                    programasvinculacioncampos = None
                    soportedelprograma=''
                    elaboradopor=''
                    perfilbeneficiarios=''
                    numerobeneficiarios=''
                    proyectosintegranprograma=''
                    perfildocentesestudiantes=''
                    planteamientoproblema=''
                    justificacion=''
                    pertinencia=''
                    objetivogeneral=''
                    objetivoespecifico=''
                    metodologia=''
                    recursos=''
                    cronograma=''
                    seguimiento=''
                    evaluacion=''
                    bibliografia=''
                    anexos=''
                    if programas.programasvinculacioncampos_set.filter(status=True).exists():
                        programasvinculacioncampos = programas.programasvinculacioncampos_set.filter(status=True)[0]
                        soportedelprograma=programasvinculacioncampos.soportedelprograma
                        elaboradopor=programasvinculacioncampos.elaboradopor
                        perfilbeneficiarios=programasvinculacioncampos.perfilbeneficiarios
                        numerobeneficiarios=programasvinculacioncampos.numerobeneficiarios
                        proyectosintegranprograma=programasvinculacioncampos.proyectosintegranprograma
                        perfildocentesestudiantes=programasvinculacioncampos.perfildocentesestudiantes
                        planteamientoproblema=programasvinculacioncampos.planteamientoproblema
                        justificacion=programasvinculacioncampos.justificacion
                        pertinencia=programasvinculacioncampos.pertinencia
                        objetivogeneral=programasvinculacioncampos.objetivogeneral
                        objetivoespecifico=programasvinculacioncampos.objetivoespecifico
                        metodologia=programasvinculacioncampos.metodologia
                        recursos=programasvinculacioncampos.recursos
                        cronograma=programasvinculacioncampos.cronograma
                        seguimiento=programasvinculacioncampos.seguimiento
                        evaluacion=programasvinculacioncampos.evaluacion
                        bibliografia=programasvinculacioncampos.bibliografia
                        anexos=programasvinculacioncampos.anexos
                    form = ProgramasVinculacionForm(initial={'nombre': programas.nombre,
                                                             'fechainicio': programas.fechainicio,
                                                             'fechaplaneado': programas.fechaplaneado,
                                                             'fechareal': programas.fechareal,
                                                             # 'lineainvestigacion': programas.lineainvestigacion,
                                                             'alcanceterritorial': programas.alcanceterritorial,
                                                             # 'areaconocimiento': programas.areaconocimiento,
                                                             # 'subareaconocimiento': programas.subareaconocimiento,
                                                             # 'subareaespecificaconocimiento': programas.subareaespecificaconocimiento,
                                                             'tiempoejecucion': programas.tiempoejecucion,
                                                             # 'sublineainvestigacion': programas.sublineainvestigacion,
                                                             'valorpresupuestointerno': programas.valorpresupuestointerno,
                                                             'valorpresupuestoexterno': programas.valorpresupuestoexterno,
                                                             'soportedelprograma': soportedelprograma,
                                                             'elaboradopor': elaboradopor,
                                                             'perfilbeneficiarios': perfilbeneficiarios,
                                                             'numerobeneficiarios': numerobeneficiarios,
                                                             'proyectosintegranprograma': proyectosintegranprograma,
                                                             'perfildocentesestudiantes': perfildocentesestudiantes,
                                                             'planteamientoproblema': planteamientoproblema,
                                                             'justificacion': justificacion,
                                                             'pertinencia': pertinencia,
                                                             'objetivogeneral': objetivogeneral,
                                                             'objetivoespecifico': objetivoespecifico,
                                                             'metodologia': metodologia,
                                                             'recursos': recursos,
                                                             'cronograma': cronograma,
                                                             'seguimiento': seguimiento,
                                                             'evaluacion': evaluacion,
                                                             'bibliografia': bibliografia,
                                                             'anexos': anexos
                                                             })
                    # form.editar(programas)
                    data['form'] = form
                    return render(request, "inv_vinculacion/editprograma.html", data)
                except Exception as ex:
                    pass

            if action == 'editproyecto':
                try:
                    data['title'] = u'Editar Proyecto'
                    data['proyectos'] = proyectos = ProyectosInvestigacion.objects.get(pk=request.GET['id'], tipo=1)

                    lista = []
                    listacant = []
                    listacarrera = []

                    form = ProyectoVinculacion1Form(initial={'programa': proyectos.programa,
                                                             'nombre': proyectos.nombre,
                                                             'tiempoejecucion': proyectos.tiempoejecucion,
                                                             'fechainicio': proyectos.fechainicio,
                                                             'fechafin': proyectos.fechafin,
                                                             'fechaPlanificacion': proyectos.fechaplaneacion,
                                                             'tipoproinstitucion': proyectos.tipoproinstitucion,
                                                             'lineainvestigacion': proyectos.lineainvestigacion,
                                                             'sublineainvestigacion': proyectos.sublineainvestigacion,
                                                             'alcanceterritorial': proyectos.alcanceterritorial,
                                                             'areaconocimiento': proyectos.areaconocimiento,
                                                             'subareaconocimiento': proyectos.subareaconocimiento,
                                                             'subareaespecificaconocimiento': proyectos.subareaespecificaconocimiento,
                                                             'valorpresupuestointerno': proyectos.valorpresupuestointerno,
                                                             'valorpresupuestoexterno': proyectos.valorpresupuestoexterno,
                                                             'presupuestototal': proyectos.presupuestototal,
                                                             # 'carreras':carreras,
                                                             'zona': [i.id for i in proyectos.zona.all()],
                                                             'canton': [i.id for i in proyectos.canton.all()],
                                                             'institucionbeneficiaria': proyectos.institucionbeneficiaria,
                                                             # 'sectorcoordenada': proyectos.sectorcoordenada,
                                                             # 'periodoejecucion': proyectos.periodoejecucion,
                                                             # 'objetivoplannacional': proyectos.objetivoplannacional,
                                                             # 'cupo':proyectos.cupo
                                                             # 'archivo': proyectos.archivo
                                                             'objetivos_PND': proyectos.objetivos_PND,
                                                             'politicas_PND': proyectos.politicas_PND,
                                                             'linea_accion': proyectos.linea_accion,
                                                             'estrategia_desarrollo': proyectos.estrategia_desarrollo,
                                                             'investigacion_institucional': proyectos.investigacion_institucional,
                                                             'necesidades_sociales': proyectos.necesidades_sociales,
                                                             'tiempo_duracion_horas': proyectos.tiempo_duracion_horas,
                                                             'distrito': proyectos.distrito,
                                                             'circuito': proyectos.circuito,
                                                             'sectorcoordenada': proyectos.sectorcoordenada,
                                                             })
                    # form.editar(proyectos)
                    data['form'] = form
                    return render(request, "inv_vinculacion/editproyecto.html", data)
                except Exception as ex:
                    pass

            if action == 'editar':
                try:
                    data['title'] = u'Proyecto de servicio comunitario'
                    # id = int(encrypt(request.GET['id']))
                    id = request.GET['id']
                    data['proyecto'] = proyecto = ProyectosInvestigacion.objects.get(id=(id))
                    # if proyecto.aprobacion == 1 or proyecto.aprobacion == 2 or proyecto.aprobacion == 4 :
                    #     return HttpResponseRedirect("/proyectovinculaciondocente?info=No puede editar la información de este proyecto.")

                    data['conv'] = FechaProyectos.objects.filter(status=True, fechainicio__lte=datetime.now(),fechafin__gte=datetime.now()).last()
                    data['beneficiarios'] = Beneficiarios.objects.filter(status=True, proyecto_id=id)
                    data['involucrados'] = Involucrado.objects.filter(status=True, proyecto_id=id)
                    data['presupuesto'] = presupuestos = Presupuesto.objects.filter(status=True, proyecto_id=id)
                    data['presupuesto2'] = Presupuesto.objects.filter(status=True, proyecto=proyecto).distinct('producto__pk')
                    total = 0.0
                    for pre in presupuestos:
                        total = total + float(pre.total)
                    data['total_presupuesto'] = total
                    data['presupuesto3'] = MarcoLogico.objects.filter(status=True, proyecto=proyecto, arbolObjetivo__tipo=2,arbolObjetivo__parentID__isnull=True).order_by('arbolObjetivo__orden')
                    subttotal = 0.0
                    for pre in Presupuesto.objects.filter(status=True, proyecto=proyecto):
                        subttotal = subttotal + float(pre.total)
                    data['total_subtotal'] = subttotal
                    data['cronograma'] = Cronograma.objects.filter(status=True, proyecto_id=id)
                    data['listaCarreras'] = CarrerasParticipantes.objects.filter(status=True, proyecto_id=id)
                    data['listaPerfilProfesional'] = PerfilProfesional.objects.filter(status=True, proyecto_id=id)
                    # data['form_cronograma'] = CronogramaForm()
                    if not ArbolProblema.objects.filter(status=True, proyecto_id=id).exists():
                        aPro = ArbolProblema(
                            proyecto=ProyectosInvestigacion.objects.get(pk=id),
                            detalle="Detalle del Problema",
                            tipo=1,
                            orden='1',
                        )
                        aPro.save(request)
                        liBase = MatrizLineaBase(
                            proyecto=ProyectosInvestigacion.objects.get(pk=id),
                            # descripcion=aPro.detalle,
                            arbolProblema=ArbolProblema.objects.get(pk=aPro.pk),
                        )
                        liBase.save(request)
                        aObj = ArbolObjetivo(
                            proyecto=ProyectosInvestigacion.objects.get(pk=id),
                            arbolProblema=ArbolProblema.objects.get(pk=aPro.pk),
                            detalle="Detalle del Proposito",
                            tipo=1,
                        )
                        aObj.save(request)
                        marLogi = MarcoLogico(
                            proyecto=ProyectosInvestigacion.objects.get(pk=id),
                            arbolObjetivo=ArbolObjetivo.objects.get(pk=aObj.pk),
                            resumen_narrativo=aObj.detalle,
                        )
                        marLogi.save(request)
                        for x in range(0,2):
                            aPro = ArbolProblema(
                                proyecto=ProyectosInvestigacion.objects.get(pk=id),
                                detalle="Detalle de la Causa"+ str(x+1),
                                tipo=2,
                                orden=x+1,
                            )
                            aPro.save(request)
                            auxIdAPro=aPro.pk
                            liBase = MatrizLineaBase(
                                proyecto=ProyectosInvestigacion.objects.get(pk=id),
                                arbolProblema=ArbolProblema.objects.get(pk=aPro.pk),
                                # descripcion=aPro.detalle,
                            )
                            liBase.save(request)
                            aObj = ArbolObjetivo(
                                proyecto=ProyectosInvestigacion.objects.get(pk=id),
                                arbolProblema=ArbolProblema.objects.get(pk=aPro.pk),
                                detalle="Detalle del Medio "+ str(x+1),
                                tipo=2,
                                orden=x + 1,
                            )
                            aObj.save(request)
                            auxIdAObj = aObj.pk
                            marLogi = MarcoLogico(
                                proyecto=ProyectosInvestigacion.objects.get(pk=id),
                                arbolObjetivo=ArbolObjetivo.objects.get(pk=aObj.pk),
                                resumen_narrativo=aObj.detalle,
                            )
                            marLogi.save(request)
                            #SUB CAUSA
                            aPro = ArbolProblema(
                                proyecto=ProyectosInvestigacion.objects.get(pk=id),
                                detalle="Detalle de la SubCausa " + str(x + 1)+'.1',
                                tipo=2,
                                parentID=ArbolProblema.objects.get(pk=auxIdAPro),
                                orden=str(x + 1) + '.1',
                            )
                            aPro.save(request)
                            liBase = MatrizLineaBase(
                                proyecto=ProyectosInvestigacion.objects.get(pk=id),
                                arbolProblema=ArbolProblema.objects.get(pk=aPro.pk),
                                # descripcion=aPro.detalle,
                            )
                            liBase.save(request)
                            aObj = ArbolObjetivo(
                                proyecto=ProyectosInvestigacion.objects.get(pk=id),
                                arbolProblema=ArbolProblema.objects.get(pk=aPro.pk),
                                detalle="Detalle del SubMedio " + str(x + 1)+'.1',
                                parentID=ArbolObjetivo.objects.get(pk=auxIdAObj),
                                tipo=2,
                                orden=str(x + 1) + '.1',
                            )
                            aObj.save(request)
                            marLogi = MarcoLogico(
                                proyecto=ProyectosInvestigacion.objects.get(pk=id),
                                arbolObjetivo=ArbolObjetivo.objects.get(pk=aObj.pk),
                                resumen_narrativo=aObj.detalle,
                            )
                            marLogi.save(request)
                        #EFECTO
                        for x in range(0,2):
                            aPro = ArbolProblema(
                                proyecto=ProyectosInvestigacion.objects.get(pk=id),
                                detalle="Detalle del Efecto " + str(x+1),
                                tipo=3,
                                orden= x+1,
                            )
                            aPro.save(request)

                            liBase = MatrizLineaBase(
                                proyecto=ProyectosInvestigacion.objects.get(pk=id),
                                arbolProblema=ArbolProblema.objects.get(pk=aPro.pk),
                                # descripcion=aPro.detalle,
                            )
                            liBase.save(request)
                            aObj = ArbolObjetivo(
                                proyecto=ProyectosInvestigacion.objects.get(pk=id),
                                arbolProblema=ArbolProblema.objects.get(pk=aPro.pk),
                                detalle="Detalle del FIN " + str(x+1),
                                tipo=3,
                                orden=x + 1,
                            )
                            aObj.save(request)
                            marLogi = MarcoLogico(
                                proyecto=ProyectosInvestigacion.objects.get(pk=id),
                                arbolObjetivo=ArbolObjetivo.objects.get(pk=aObj.pk),
                                resumen_narrativo=aObj.detalle,
                            )
                            marLogi.save(request)

                    data['aPro_problema']= ArbolProblema.objects.get(status=True, proyecto_id=id, tipo=1)
                    data['aPro_causa'] = ArbolProblema.objects.filter(status=True, proyecto_id=id, tipo=2).order_by('orden')
                    data['aPro_efecto'] = ArbolProblema.objects.filter(status=True, proyecto_id=id, tipo=3).order_by('orden')
                    data['aObj_proposito'] = ArbolObjetivo.objects.get(status=True, proyecto_id=id, tipo=1)
                    data['aObj_medio'] = ArbolObjetivo.objects.filter(status=True, proyecto_id=id, tipo=2).order_by('orden')
                    data['aObj_fin'] = ArbolObjetivo.objects.filter(status=True, proyecto_id=id, tipo=3).order_by('orden')
                    efectos = MatrizLineaBase.objects.filter(status=True, proyecto_id=id,arbolProblema__tipo=3).order_by('arbolProblema__orden')
                    causas= MatrizLineaBase.objects.filter(status=True, proyecto_id=id,arbolProblema__tipo=2).order_by('arbolProblema__orden')
                    problema= MatrizLineaBase.objects.filter(status=True, proyecto_id=id,arbolProblema__tipo=1)
                    data['aPro_lineaBase'] = list(efectos)+list(problema)+list(causas)
                    data['aPro_marcoLogico_fin'] = MarcoLogico.objects.filter(status=True, proyecto_id=id, arbolObjetivo__tipo=3)
                    data['aPro_marcoLogico_proposito'] = MarcoLogico.objects.filter(status=True, proyecto_id=id, arbolObjetivo__tipo=1)
                    data['aPro_marcoLogico_componentes'] = MarcoLogico.objects.filter(status=True, proyecto_id=id, arbolObjetivo__tipo=2, arbolObjetivo__parentID__isnull = True).order_by('arbolObjetivo__orden')
                    # arbolObjetivo__parentID__isnull = True
                    data['aPro_marcoLogico_acciones'] = acciones = MarcoLogico.objects.filter(status=True, proyecto_id=id, arbolObjetivo__tipo=2, arbolObjetivo__parentID__isnull = False).order_by('arbolObjetivo__orden')
                    # arbolObjetivo__parentID__isnull = False
                    if Problema.objects.filter(status=True, proyecto_id=id).exists():
                        data['redaccion'] = Problema.objects.get(status=True, proyecto_id=id)
                    data['lista']=ParticipantesMatrices.objects.filter(status=True, proyecto_id=id,tipoparticipante__tipo=1).order_by('fecha_creacion')
                    data['datoSecundario']=DatoSecundario.objects.filter(status=True, proyecto_id=id)
                    data['anexos']= Anexos.objects.filter(status=True, proyecto_id=id)
                    #control de pestañas
                    #
                    proyecto = ProyectosInvestigacion.objects.filter(pk=id).count()
                    if proyecto >= 1:
                        data['bool_proyecto'] = True

                    carrerasParticipantes = CarrerasParticipantes.objects.filter(proyecto_id=id, status=True)
                    carrera = False
                    for x in carrerasParticipantes:
                        perfil = PerfilProfesional.objects.filter(carreras_participantes=x, status = True).exists()
                        if not perfil:
                            carrera = False
                            break
                        else:
                            carrera =True

                    if carrera:
                        data['bool_car_participantes'] = True

                    lider = ParticipantesMatrices.objects.filter(proyecto_id=id, tipoparticipante__id = 1,status=True).count()
                    promotor = ParticipantesMatrices.objects.filter(proyecto_id=id, tipoparticipante__id = 2, status=True).count()
                    if lider == 1 and promotor>=1:
                        data['bool_doc_participantes'] = True

                    beneficiario = Beneficiarios.objects.filter(proyecto_id=(id), status=True).count()
                    if beneficiario >= 1:
                        data['bool_beneficiarios'] = True

                    linea = ProyectosInvestigacion.objects.get(pk=(id))
                    if linea.linea_base:
                        lineaBase = MatrizLineaBase.objects.filter(proyecto_id=(id), editado=False, status=True).count()
                        if not lineaBase >= 1:
                            data['bool_linea'] = True
                    else:
                        if DatoSecundario.objects.filter(status=True,proyecto=linea).count() >= 1:
                            data['bool_linea'] = True
                    marcoLogico = MarcoLogico.objects.filter(proyecto_id=(id), editado=False, status=True).count()
                    component = MarcoLogico.objects.filter(status=True, proyecto_id=id, arbolObjetivo__tipo=2,arbolObjetivo__parentID__isnull=True)
                    porcentaje=0.0
                    for comp in component:
                        porcentaje = porcentaje + float(comp.cumplimiento)
                    if not marcoLogico >= 1:
                        if porcentaje == 100.0:
                            data['bool_marcoLogico'] = True

                    arbolObjetivo = ArbolObjetivo.objects.filter(proyecto_id=(id), editado=False, status=True).count()
                    if not arbolObjetivo >= 1:
                        data['bool_arbolObjetivo'] = True

                    arbolProblema = ArbolProblema.objects.filter(proyecto_id=(id), editado=False, status=True).count()
                    if not arbolProblema >= 1:
                        data['bool_arbolProblema'] = True

                    for x in acciones:
                        crono = Cronograma.objects.filter(aobjetivo=x.arbolObjetivo, status = True).exists()
                        if not crono:
                            cronograma = False
                            break
                        else:
                            cronograma =True

                    if cronograma:
                        data['bool_cronograma'] = True

                    presupuesto = Presupuesto.objects.filter(proyecto_id=(id), status= True).count()
                    if presupuesto >= 1:
                        data['bool_presupuesto'] = True

                    anexos = Anexos.objects.filter(proyecto_id=(id), status=True).count()
                    if anexos >= 3:
                        data['bool_anexos'] = True

                    return render(request, "inv_vinculacion/crearproyecto.html", data)
                except Exception as ex:
                    pass

            if action == 'generarpdf':
                try:
                    data['proyecto'] = proyecto = ProyectosInvestigacion.objects.get(id=request.GET['id'])
                    data['carrerasparticipantes'] = CarrerasParticipantes.objects.filter(status=True, proyecto=proyecto)
                    data['presupuesto'] = Presupuesto.objects.filter(status=True, proyecto=proyecto)
                    data['presupuesto3'] = MarcoLogico.objects.filter(status=True, proyecto=proyecto,arbolObjetivo__tipo=2,arbolObjetivo__parentID__isnull=True).order_by('arbolObjetivo__orden')
                    subttotal = 0.0
                    for pre in Presupuesto.objects.filter(status=True, proyecto=proyecto):
                        subttotal = subttotal + float(pre.total)
                    data['total_subtotal'] = subttotal
                    data['docentesparticipantes'] = docentes = ParticipantesMatrices.objects.filter(status=True, proyecto=proyecto,tipoparticipante__tipo=1).order_by('fecha_creacion')
                    data['perfilprofesional'] = PerfilProfesional.objects.filter(status=True, proyecto=proyecto)
                    data['beneficiarios'] = Beneficiarios.objects.filter(status=True, proyecto=proyecto)
                    data['problema'] = Problema.objects.get(status=True, proyecto=proyecto)
                    data['marcologico'] = MarcoLogico.objects.filter(status=True, proyecto=proyecto)
                    data['aPro_marcoLogico_fin'] = MarcoLogico.objects.filter(status=True, proyecto=proyecto,arbolObjetivo__tipo=3)
                    data['aPro_marcoLogico_proposito'] = MarcoLogico.objects.filter(status=True, proyecto=proyecto,arbolObjetivo__tipo=1)
                    data['aPro_marcoLogico_componentes'] = MarcoLogico.objects.filter(status=True, proyecto=proyecto,arbolObjetivo__tipo=2,arbolObjetivo__parentID__isnull=True).order_by('arbolObjetivo__orden')
                    data['aPro_marcoLogico_acciones'] = acciones = MarcoLogico.objects.filter(status=True,proyecto=proyecto,arbolObjetivo__tipo=2,arbolObjetivo__parentID__isnull=False).order_by('arbolObjetivo__orden')
                    data['presupuesto2'] = Presupuesto.objects.filter(status=True, proyecto=proyecto).distinct('producto__pk')
                    total = 0.0
                    for pre in Presupuesto.objects.filter(status=True, proyecto=proyecto):
                        total = total + float(pre.total)
                    data['total_presupuesto'] = total

                    data['efectos'] = ArbolProblema.objects.filter(status=True, proyecto=proyecto, tipo=3).order_by('orden')
                    data['causas'] = ArbolProblema.objects.filter(status=True, proyecto=proyecto, tipo=2).order_by('orden')
                    data['problemaarbol'] = ArbolProblema.objects.filter(status=True, proyecto=proyecto, tipo=1).order_by('orden')
                    data['aObj_proposito'] = ArbolObjetivo.objects.filter(status=True, proyecto=proyecto,tipo=1).order_by('orden')
                    data['aObj_medio'] = ArbolObjetivo.objects.filter(status=True, proyecto=proyecto,tipo=2).order_by('orden')
                    data['aObj_fin'] = ArbolObjetivo.objects.filter(status=True, proyecto=proyecto, tipo=3).order_by('orden')
                    data['datosecundario'] = DatoSecundario.objects.filter(status=True, proyecto=proyecto)
                    data['efec'] = MatrizLineaBase.objects.filter(status=True, proyecto=proyecto,arbolProblema__tipo=3).order_by('arbolProblema__orden')
                    data['caus'] = MatrizLineaBase.objects.filter(status=True, proyecto=proyecto,arbolProblema__tipo=2).order_by('arbolProblema__orden')
                    data['prob'] = MatrizLineaBase.objects.filter(status=True, proyecto=proyecto,arbolProblema__tipo=1)
                    data['aPro_lineaBase'] = list(data['efec']) + list(data['prob']) + list(data['caus'])
                    data['lineabase'] = MatrizLineaBase.objects.filter(status=True, proyecto=proyecto)
                    data['cronograma'] = Cronograma.objects.filter(status=True, proyecto=proyecto)
                    data['fecha'] = datetime.now().strftime("%Y-%m-%d")
                    # data['periodo'] = periodo = Periodo.objects.filter(status=True, tipo__pk=2,inicio__gte=proyecto.fecha_creacion).first()
                    data['periodo'] = periodo
                    lider = ParticipantesMatrices.objects.get(status=True, proyecto=proyecto, tipoparticipante__pk=1)
                    carrera_lider = lider.profesor.carrera_comun_periodo(periodo)
                    if carrera_lider:
                        coordinacion = carrera_lider.coordinacion_carrera()
                        decano = coordinacion.responsable_periododos(periodo,1)
                        data['coordinacion'] = coordinacion
                        data['decano'] = decano
                    data['lider'] = lider.profesor.persona
                    data['profesor'] = lider.profesor

                    return conviert_html_to_pdf(
                        'proyectovinculaciondocente/proyectovinculacion.html',
                        {
                            'pagesize': 'A4',
                            'data': data,
                        }
                    )
                except Exception as ex:
                    print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))



            elif action == 'aprobacionproyecto':
                try:
                    data['title'] = u'Aprobacion de Proyecto'
                    data['proyectos'] = proyectos = ProyectosInvestigacion.objects.get(pk=request.GET['id'], tipo=1)
                    form = AprobacionProyecto()
                    data['form'] = form
                    return render(request, "inv_vinculacion/aprobacionproyecto.html", data)
                except Exception as ex:
                    pass

            elif action == 'listafechas':
                try:
                    data['title'] = u'Lista de Fechas de Convocatorias'
                    data['fechas'] = fecha = FechaProyectos.objects.filter(status=True).order_by('-fecha_creacion')
                    return render(request, "inv_vinculacion/listafechas.html", data)
                except Exception as ex:
                    pass

            elif action == 'listaresolucion':
                try:
                    data['title'] = u'Lista de resolución'
                    data['resolucion'] = fecha = Resolucion.objects.filter(status=True, convocatoria__id= request.GET['conv']).order_by('-fecha_creacion')
                    data['conv'] = request.GET['conv']
                    return render(request, "inv_vinculacion/listaresolucion.html", data)
                except Exception as ex:
                    pass

            elif action == 'viewcargobeneficiario':
                try:
                    data['title'] = u'Lista de cargo de beneficiarios'
                    data['fechas'] = fecha = CargoBeneficiario.objects.filter(status=True).order_by('-fecha_creacion')
                    return render(request, "inv_vinculacion/viewcargobeneficario.html", data)
                except Exception as ex:
                    pass

            elif action == 'listaInsumos':
                try:
                    data['title'] = u'Lista de insumos para proyectos'
                    search = None
                    ids = None
                    filtros = Q(status=True)
                    if 's' in request.GET:
                        search = request.GET['s']
                        filtros = filtros & Q(rubro__icontains=search) | Q(especificacion__icontains=search)
                        if ' ' in search:
                            s = search.split(" ")
                    suministro = Suministro.objects.filter(filtros).order_by('-fecha_creacion')
                    paging = MiPaginador(suministro, 25)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['page'] = page
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['insumos'] = page.object_list
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""

                    return render(request, "inv_vinculacion/viewinsumos.html", data)
                except Exception as ex:
                    pass

            elif action == 'addevidenciasprogramas':
                try:
                    data['title'] = u'Evidencia Programa'
                    data['form'] = EvidenciaForm
                    data['id'] = request.GET['id']
                    data['idevidencia'] = request.GET['idevidencia']
                    template = get_template("inv_vinculacion/add_evidencia.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content})
                except Exception as ex:
                    pass

            elif action == 'evidenciasprogramas':
                try:
                    data['title'] = u'Evidencia Programa'
                    data['programas'] = programas = ProgramasInvestigacion.objects.get(pk=request.GET['id'])
                    data['evidencias'] = Evidencia.objects.filter(status=True, matrizevidencia_id=1)
                    data['formevidencias'] = EvidenciaForm()
                    return render(request, "inv_vinculacion/evidenciasprogramas.html", data)
                except Exception as ex:
                    pass

            elif action == 'deleteprograma':
                try:
                    data['title'] = u'Eliminar Programa'
                    data['programa'] = ProgramasInvestigacion.objects.get(pk=request.GET['id'])
                    return render(request, "inv_vinculacion/deleteprograma.html", data)
                except Exception as ex:
                    pass

            elif action == 'deleteparticipanteproyecto':
                try:
                    data['title'] = u'Eliminar Participante'
                    tipo = request.GET['tipo']
                    data['participante'] = participante = ParticipantesMatrices.objects.get(pk=request.GET['id'])
                    if tipo == '1':
                        data['nombres'] = participante.profesor.persona.nombre_completo() if participante.profesor else participante.externo.persona.nombre_completo()
                        return render(request, "inv_vinculacion/deleteparticipanteproyecto.html", data)
                    if tipo == '2':
                        data['nombres'] = participante.inscripcion.persona.nombre_completo()
                        return render(request, "inv_vinculacion/deleteparticipanteproyecto_alu.html", data)
                    if tipo == '3':
                        data['nombres'] = participante.administrativo.persona.nombre_completo()
                        return render(request, "inv_vinculacion/deleteparticipanteproyecto.html", data)
                except Exception as ex:
                    pass

            elif action == 'deletepresupuestoproyecto':
                try:
                    data['title'] = u'Eliminar Presupuesto'
                    data['presupuesto'] = PresupuestosProyecto.objects.get(pk=request.GET['id'])
                    return render(request, "inv_vinculacion/deletepresupuestoproyecto.html", data)
                except Exception as ex:
                    pass

            elif action == 'listadoproyectos':
                try:
                    filtros = Q(status=True, tipo = 1)
                    data['title'] = u'Listado de Proyectos'
                    data['estadoproyectos'] = ESTADOS_PROYECTO_VINCULACION_INVESTIGACION
                    search = None
                    ids = None
                    tipobus = None
                    inscripcionid = None
                    if 'id' in request.GET:
                        ids = request.GET['id']
                        programasinvestigacion = ProyectosInvestigacion.objects.select_related().filter(pk=ids,tipo=1).order_by('nombre')
                    elif 'inscripcionid' in request.GET:
                        inscripcionid = request.GET['inscripcionid']
                        programasinvestigacion = Graduado.objects.filter(inscripcion__carrera__in=miscarreras).filter(inscripcion__id=inscripcionid)

                    if 's' in request.GET:
                        search = request.GET['s']
                        if search.isdigit():
                            filtros = filtros & Q(fechainicio__year=search) | Q(id=search)
                        else:
                            filtros = filtros & Q(nombre__icontains=search)
                    if 'estado' in request.GET:
                        estado = request.GET['estado']
                        data['estado'] = int(estado)
                        filtros = filtros & Q(aprobacion=estado)

                    programasinvestigacion = ProyectosInvestigacion.objects.select_related().filter(filtros).order_by('-aprobacion','-programa__fecha_creacion', 'nombre')
                    paging = MiPaginador(programasinvestigacion, 25)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['page'] = page
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['proyectos'] = page.object_list
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""

                    conv = FechaProyectos.objects.filter(status=True, fechainicio__lte=datetime.now(),fechafin__gte=datetime.now())
                    if conv.exists():
                        data['convocatoria'] = conv.last()
                        if not ProyectosInvestigacion.objects.filter(status=True, convocatoria=conv.last(),usuario_creacion=request.user).exists():
                            data['conv_activa'] = True

                    return render(request, "inv_vinculacion/listaproyectos.html", data)
                except Exception as ex:
                    pass


            elif action == 'ejecucion':
                try:
                    est, search, desde, hasta, mdesde, mhasta, filtros, url = \
                        request.GET.get('est', ''), \
                        request.GET.get('s', ''), \
                        request.GET.get('desde', ''), \
                        request.GET.get('hasta', ''), \
                        request.GET.get('mdesde', ''), \
                        request.GET.get('mhasta', ''), \
                        Q(status=True), \
                        ''
                    if est:
                        data['ejec_estado'] = estado = int(est)
                        url += "&est={}".format(estado)
                        if estado == 1:
                            filtros = filtros & Q(estado_finalizado=False)
                        elif estado == 2:
                            filtros = filtros & Q(estado_finalizado=True)
                    if desde and hasta:
                        data['ejec_desde'] = desde
                        data['ejec_hasta'] = hasta
                        url += "&desde={}".format(desde)
                        url += "&hasta={}".format(hasta)
                        filtros = filtros & (Q(fecha_inicio__gte=desde) & Q(fecha_fin__lte=hasta))

                    elif desde:
                        data['ejec_desde'] = desde
                        url += "&desde={}".format(desde)
                        filtros = filtros & Q(fecha_inicio__gte=desde)
                    elif hasta:
                        data['ejec_hasta'] = hasta
                        url += "&hasta={}".format(hasta)
                        filtros = filtros & Q(fecha_fin__lte=hasta)

                    if mdesde and mhasta:
                        data['ejec_mes_desde'] = mdesde
                        data['ejec_mes_hasta'] = mhasta
                        url += "&mdesde={}".format(mdesde)
                        url += "&mhasta={}".format(mhasta)
                        filtros = filtros & (Q(fecha_inicio__month__gte=datetime.strptime(mdesde, '%Y-%m').date().month) &
                                             Q(fecha_fin__month__lte=datetime.strptime(mhasta, '%Y-%m').date().month)&
                                             Q(fecha_inicio__year=datetime.strptime(mdesde, '%Y-%m').date().year) &
                                             Q(fecha_fin__year=datetime.strptime(mhasta, '%Y-%m').date().year))

                    elif mdesde:
                        data['ejec_mes_desde'] = mdesde
                        url += "&mdesde={}".format(mdesde)
                        filtros = filtros & (Q(fecha_inicio__month__lte=datetime.strptime(mdesde, '%Y-%m').date().month) &
                                             Q(fecha_fin__month__gte=datetime.strptime(mdesde, '%Y-%m').date().month) &
                                             Q(fecha_inicio__year=datetime.strptime(mdesde, '%Y-%m').date().year)

                                             )

                    elif mhasta:
                        data['ejec_mes_hasta'] = mhasta
                        url += "&mhasta={}".format(mhasta)
                        filtros = filtros & (Q(fecha_fin__month__lte=datetime.strptime(mhasta, '%Y-%m').date().month) &
                                             Q(fecha_fin__year=datetime.strptime(mhasta, '%Y-%m').date().year))


                    if search:
                        data['search'] = search
                        s = search.split()
                        if len(s) == 1:
                            filtros = filtros & (Q(responsable__persona__apellido1__icontains=search) |
                                                 Q(responsable__persona__apellido2__icontains=search) |
                                                 Q(responsable__persona__cedula__icontains=search) |
                                                 Q(responsable__persona__nombres__icontains=search))
                            url += '&search={}'.format(search)
                        else:
                            filtros = filtros & (Q(responsable__persona__apellido1__icontains=s[0]) &
                                                 Q(responsable__persona__apellido2__icontains=s[1]))
                            url += '&search={}'.format(search)
                    data['url'] = url

                    data['title'] = u'Cronograma de ejecución'
                    data['proyecto'] = proyecto = ProyectosInvestigacion.objects.get(pk=request.GET['id'],tipo=1)

                    data['cronograma'] = cronograma = Cronograma.objects.filter(filtros, proyecto=proyecto)
                    data['aPro_marcoLogico_acciones'] = MarcoLogico.objects.filter(status=True, proyecto=proyecto,arbolObjetivo__tipo=2,arbolObjetivo__parentID__isnull=False)
                    data['aPro_marcoLogico_componentes'] = MarcoLogico.objects.filter(status=True, proyecto=proyecto,arbolObjetivo__tipo=2,arbolObjetivo__parentID__isnull=True).order_by('arbolObjetivo__orden')
                    data['total_pendientes'] = cronograma.filter(estado_finalizado=False).count()
                    data['total_finalizados'] = cronograma.filter(estado_finalizado=True).count()
                    return render(request, "inv_vinculacion/adm_ejecucion.html", data)
                except Exception as ex:
                    pass

            elif action == 'veravance':
                try:
                    data['title'] = u'Ver avance'
                    data['id'] = id = request.GET['id']
                    data['tarea'] = tarea = Cronograma.objects.get(pk=id)
                    data['avance'] = DetalleCumplimiento.objects.filter(status=True,tarea= tarea ).order_by('-fecha_ingreso')

                    ESTADOS_INFORMES = (
                        (1, u'APROBAR'),
                        (2, u'RECHAZAR')
                    )

                    data['estados_informes'] = ESTADOS_INFORMES

                    template = get_template("inv_vinculacion/adm_ver_avance.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content})
                except Exception as ex:
                    pass

            elif action == 'aprobacion_informe':
                try:
                    data['title'] = u'Aprobacion de Informe'
                    data['id'] = id = request.GET['id']
                    data['proyecto']  = DetalleCumplimiento.objects.get(pk=id)
                    data['form'] = AprobacionCumplimientoForm()
                    template = get_template("inv_vinculacion/aprobacion_adm.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content})
                except Exception as ex:
                    pass


            elif action == 'evidenciasproyectos':
                try:
                    data['title'] = u'Evidencia Proyectos'
                    data['proyectos'] = proyectosinvestigacion = ProyectosInvestigacion.objects.get(pk=request.GET['id'],tipo=1)
                    data['evidencias'] = Evidencia.objects.filter(status=True, matrizevidencia_id=2)
                    data['formevidencias'] = EvidenciaForm()

                    data['informemarcologicoproyectosinvestigaciones'] = InformeMarcoLogicoProyectosInvestigacion.objects.filter(proyectovinculacion=proyectosinvestigacion, status=True).order_by('fecha')
                    return render(request, "inv_vinculacion/evidenciasproyectos.html", data)
                except Exception as ex:
                    pass

            elif action == 'addevidenciasproyectos':
                try:
                    data['title'] = u'Evidencia Programa'
                    data['form'] = EvidenciaForm
                    data['id'] = request.GET['id']
                    data['idevidencia'] = request.GET['idevidencia']
                    template = get_template("inv_vinculacion/add_evidenciaproyectos.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content})
                except Exception as ex:
                    pass

            elif action == 'addfecha':
                try:
                    data['title'] = u'Rango para subir proyecto'
                    data['form'] = FechaProyectosFrom

                    template = get_template("inv_vinculacion/addfecha.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content})
                except Exception as ex:
                    pass

            elif action == 'addfechaentrega':
                try:
                    titulo = 'Definir fecha de entrega'
                    data['form'] = FechaEntregaProyectoFrom
                    data['idproyecto'] = request.GET['id']

                    template = get_template("inv_vinculacion/addfechaentrega.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok",'titulo':titulo, 'data': json_content})
                except Exception as ex:
                    pass

            elif action == 'actualizaConvocatoria':
                try:
                    titulo = 'Convocatoria'
                    proy = ProyectosInvestigacion.objects.get(id=request.GET['id'])
                    data['form'] = CambioConvocatoriaForm(initial={
                        'convocatoria': proy.convocatoria
                    })
                    data['idproyecto'] = request.GET['id']

                    template = get_template("inv_vinculacion/editConvocatoria.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok",'titulo':titulo, 'data': json_content})
                except Exception as ex:
                    pass

            elif action == 'addfechaaprobacion':
                try:
                    titulo = 'Definir fecha de aprobación'
                    data['form'] = FechaEntregaProyectoFrom
                    data['idproyecto'] = request.GET['id']

                    template = get_template("inv_vinculacion/addfechaaprobacion.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", "titulo": titulo, 'data': json_content})
                except Exception as ex:
                    pass

            elif action == 'addresolucion':
                try:
                    data['title'] = u'Adicionar beneficiarios'
                    data['form'] = ResolucionForm()
                    data['conv'] = request.GET['conv']
                    template = get_template("inv_vinculacion/addresolucion.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content})
                except Exception as ex:
                    pass

            elif action == 'editresolucion':
                try:
                    data['title'] = u'Editar beneficiarios'
                    data['resolucion']=datos = Resolucion.objects.get(pk=int(request.GET['id']))
                    data['form'] = ResolucionForm(initial={
                        'resolucion': datos.resolucion,
                    })
                    template = get_template("inv_vinculacion/editresolucion.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content})
                except Exception as ex:
                    pass

            elif action == 'deleteresolucion':
                try:
                    data['title'] = u'Eliminar resolución'
                    data['resolucion'] =  Resolucion.objects.get(pk=request.GET['id'])
                    return render(request, "inv_vinculacion/deleteresolucion.html", data)
                except Exception as ex:
                    pass


            elif action == 'addcargobeneficiario':
                try:
                    data['title'] = u'Crear cargo beneficiario'
                    data['form'] = CargoBeneficiarioFrom

                    template = get_template("inv_vinculacion/addcargobeneficiario.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content})
                except Exception as ex:
                    pass

            elif action == 'editfecha':
                try:
                    data['title'] = u'Fecha de convocatoria'
                    convocatoria = FechaProyectos.objects.get(pk=request.GET['id'])
                    data['form'] = FechaProyectosFrom(initial={
                                                    'descripcion': convocatoria.descripcion,
                                                    'fechainicio': convocatoria.fechainicio,
                                                    'fechafin': convocatoria.fechafin,
                                                    })
                    template = get_template("inv_vinculacion/editfecha.html")
                    data['convocatoria'] = convocatoria
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content})
                except Exception as ex:
                    pass

            elif action == 'editcargobeneficiario':
                try:
                    data['title'] = u'Editrar cargo'
                    datos = CargoBeneficiario.objects.get(pk=request.GET['id'])
                    data['form'] = CargoBeneficiarioFrom(initial={
                                                    'descripcion': datos.descripcion,
                                                    })
                    template = get_template("inv_vinculacion/editcargobeneficiario.html")
                    data['convocatoria'] = datos
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content})
                except Exception as ex:
                    pass

            elif action == 'verproyecto':
                try:
                    data['title'] = u'Proyecto de servicio comunitario'
                    # id = int(encrypt(request.GET['id']))
                    id = request.GET['id']
                    data['proyecto'] = proyecto = ProyectosInvestigacion.objects.get(id=(id))

                    data['conv'] = FechaProyectos.objects.filter(status=True, fechainicio__lte=datetime.now(),
                                                                 fechafin__gte=datetime.now()).last()
                    data['beneficiarios'] = Beneficiarios.objects.filter(status=True, proyecto_id=id)
                    data['involucrados'] = Involucrado.objects.filter(status=True, proyecto_id=id)
                    data['presupuesto'] = Presupuesto.objects.filter(status=True, proyecto_id=id)
                    data['presupuesto2'] = Presupuesto.objects.filter(status=True, proyecto=proyecto).distinct(
                        'suministro__pk')
                    total = 0.0
                    for pre in Presupuesto.objects.filter(status=True, proyecto=proyecto):
                        total = total + float(pre.total)
                    data['total_presupuesto'] = total
                    data['presupuesto3'] = MarcoLogico.objects.filter(status=True, proyecto=proyecto,
                                                                      arbolObjetivo__tipo=2,
                                                                      arbolObjetivo__parentID__isnull=True).order_by(
                        'arbolObjetivo__orden')
                    subttotal = 0.0
                    for pre in Presupuesto.objects.filter(status=True, proyecto=proyecto):
                        subttotal = subttotal + float(pre.total)
                    data['total_subtotal'] = subttotal
                    data['cronograma'] = Cronograma.objects.filter(status=True, proyecto_id=id)
                    data['listaCarreras'] = CarrerasParticipantes.objects.filter(status=True, proyecto_id=id)
                    data['listaPerfilProfesional'] = PerfilProfesional.objects.filter(status=True, proyecto_id=id)
                    if ArbolProblema.objects.filter(status=True, proyecto_id=id, tipo=1).exists():
                        data['aPro_problema'] = ArbolProblema.objects.get(status=True, proyecto_id=id, tipo=1)

                    data['aPro_causa'] = ArbolProblema.objects.filter(status=True, proyecto_id=id, tipo=2).order_by(
                        'orden')
                    data['aPro_efecto'] = ArbolProblema.objects.filter(status=True, proyecto_id=id, tipo=3).order_by(
                        'orden')
                    if ArbolObjetivo.objects.filter(status=True, proyecto_id=id, tipo=1).exists():
                        data['aObj_proposito'] = ArbolObjetivo.objects.get(status=True, proyecto_id=id, tipo=1)
                    data['aObj_medio'] = ArbolObjetivo.objects.filter(status=True, proyecto_id=id, tipo=2).order_by(
                        'orden')
                    data['aObj_fin'] = ArbolObjetivo.objects.filter(status=True, proyecto_id=id, tipo=3).order_by(
                        'orden')
                    efectos = MatrizLineaBase.objects.filter(status=True, proyecto_id=id,
                                                             arbolProblema__tipo=3).order_by(
                        'arbolProblema__orden')
                    causas = MatrizLineaBase.objects.filter(status=True, proyecto_id=id,
                                                            arbolProblema__tipo=2).order_by(
                        'arbolProblema__orden')
                    problema = MatrizLineaBase.objects.filter(status=True, proyecto_id=id, arbolProblema__tipo=1)
                    data['aPro_lineaBase'] = list(efectos) + list(problema) + list(causas)
                    data['aPro_marcoLogico_fin'] = MarcoLogico.objects.filter(status=True, proyecto_id=id,
                                                                              arbolObjetivo__tipo=3)
                    data['aPro_marcoLogico_proposito'] = MarcoLogico.objects.filter(status=True, proyecto_id=id,
                                                                                    arbolObjetivo__tipo=1)
                    data['aPro_marcoLogico_componentes'] = MarcoLogico.objects.filter(status=True, proyecto_id=id,
                                                                                      arbolObjetivo__tipo=2,
                                                                                      arbolObjetivo__parentID__isnull=True).order_by(
                        'arbolObjetivo__orden')
                    data['aPro_marcoLogico_acciones'] = acciones = MarcoLogico.objects.filter(status=True,
                                                                                              proyecto_id=id,
                                                                                              arbolObjetivo__tipo=2,
                                                                                              arbolObjetivo__parentID__isnull=False).order_by(
                        'arbolObjetivo__orden')
                    if Problema.objects.filter(status=True, proyecto_id=id).exists():
                        data['redaccion'] = Problema.objects.get(status=True, proyecto_id=id)
                    data['lista'] = ParticipantesMatrices.objects.filter(status=True, proyecto_id=id,tipoparticipante__tipo=1).order_by('fecha_creacion')
                    data['datoSecundario'] = DatoSecundario.objects.filter(status=True, proyecto_id=id)
                    data['anexos'] = Anexos.objects.filter(status=True, proyecto_id=id)
                    # control de pestañas
                    #
                    proyecto = ProyectosInvestigacion.objects.filter(pk=id).count()
                    if proyecto >= 1:
                        data['bool_proyecto'] = True

                    carrerasParticipantes = CarrerasParticipantes.objects.filter(proyecto_id=id, status=True)
                    carrera = False
                    for x in carrerasParticipantes:
                        perfil = PerfilProfesional.objects.filter(carreras_participantes=x, status=True).exists()
                        if not perfil:
                            carrera = False
                            break
                        else:
                            carrera = True

                    if carrera:
                        data['bool_car_participantes'] = True

                    lider = ParticipantesMatrices.objects.filter(proyecto_id=id, tipoparticipante__id=1,
                                                                 status=True).count()
                    promotor = ParticipantesMatrices.objects.filter(proyecto_id=(id), tipoparticipante__id=2,
                                                                    status=True).count()
                    if lider == 1 and promotor >= 1:
                        data['bool_doc_participantes'] = True

                    beneficiario = Beneficiarios.objects.filter(proyecto_id=(id), status=True).count()
                    if beneficiario >= 1:
                        data['bool_beneficiarios'] = True

                    linea = ProyectosInvestigacion.objects.get(pk=(id))
                    if linea.linea_base:
                        lineaBase = MatrizLineaBase.objects.filter(proyecto_id=(id), editado=False, status=True).count()
                        if not lineaBase >= 1:
                            data['bool_linea'] = True
                    else:
                        if DatoSecundario.objects.filter(status=True, proyecto=linea).count() >= 1:
                            data['bool_linea'] = True
                    marcoLogico = MarcoLogico.objects.filter(proyecto_id=(id), editado=False, status=True).count()
                    if not marcoLogico >= 1:
                        data['bool_marcoLogico'] = True

                    arbolObjetivo = ArbolObjetivo.objects.filter(proyecto_id=(id), editado=False, status=True).count()
                    if not arbolObjetivo >= 1:
                        data['bool_arbolObjetivo'] = True

                    arbolProblema = ArbolProblema.objects.filter(proyecto_id=(id), editado=False, status=True).count()
                    if not arbolProblema >= 1:
                        data['bool_arbolProblema'] = True
                    cronograma = False
                    for x in acciones:
                        crono = Cronograma.objects.filter(aobjetivo=x.arbolObjetivo, status=True).exists()
                        if not crono:
                            cronograma = False
                            break
                        else:
                            cronograma = True

                    if cronograma:
                        data['bool_cronograma'] = True

                    presupuesto = Presupuesto.objects.filter(proyecto_id=(id), status=True).count()
                    if presupuesto >= 1:
                        data['bool_presupuesto'] = True

                    anexos = Anexos.objects.filter(proyecto_id=(id), status=True).count()
                    if anexos >= 3:
                        data['bool_anexos'] = True

                    return render(request, "inv_vinculacion/crearproyecto.html", data)
                except Exception as ex:
                    print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
                    pass

            elif action == 'addfechafin':
                try:
                    data['title'] = u'Subir Fecha del proyecto'
                    data['form'] = FechaFinProyectosFrom()

                    template = get_template("inv_vinculacion/addfecha.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content})
                except Exception as ex:
                    pass

            elif action == 'addparticipantesdocentes':
                try:
                    data['title'] = u'Participante Docente'
                    form = ParticipanteProfesorForm()
                    var = request.GET['tipoparticipante']
                    form.adicionar(var)
                    data['form'] = form
                    data['id'] = request.GET['idproyecto']
                    return render(request, "inv_vinculacion/addparticipantedocente.html", data)
                except Exception as ex:
                    pass

            if action == 'addparticipantesdocentesp':
                try:
                    data['title'] = u'Agregar Participante Docente'
                    form = ParticipanteProfesorVinculacionForm()
                    data['form'] = form
                    data['id'] = id = request.GET['idproyecto']
                    return render(request, "inv_vinculacion/addparticipantedocentepvinc.html", data)
                except Exception as ex:
                    pass

            if action == 'editparticipantesdocentes':
                try:
                    data['title'] = u'Editar Participante Docente'
                    data['id'] = id = request.GET['id']

                    docente = ParticipantesMatrices.objects.get(pk=id)
                    data['idproyecto'] = docente.proyecto.pk
                    data['docente'] = docente

                    form = ParticipanteProfesorVinculacionForm(
                        initial={
                            'tipo': 1 if docente.profesor else 2,
                            # 'profesor': docente.profesor,
                            'horas': docente.horas,
                            'nivel': [i.id for i in docente.nivel.all()]
                        }
                    )

                    data['form'] = form
                    data['accionbuscar'] = 'Profesor' if docente.profesor else 'Persona'
                    data['tipodocente'] = 1 if docente.profesor else 2
                    data['iddocente'] = docente.profesor.id if docente.profesor else docente.externo.persona.id
                    data['nomdocente'] = docente.profesor.persona.flexbox_repr() if docente.profesor else docente.externo.persona.flexbox_repr()

                    return render(request, "inv_vinculacion/editparticipantedocentepvinc.html", data)
                except Exception as ex:
                    pass

            elif action == 'addparticipantesestudiantes':
                try:
                    data['title'] = u'Participante Estudiante'
                    data['form'] = ParticipanteEstudianteForm
                    data['id'] = request.GET['idproyecto']
                    return render(request, "inv_vinculacion/addparticipanteestudiante.html", data)
                except Exception as ex:
                    pass

            elif action == 'addparticipantesadministrativos':
                try:
                    data['title'] = u'Participante Administrativo'
                    data['form'] = ParticipanteAdministrativoForm
                    data['id'] = request.GET['idproyecto']
                    return render(request, "inv_vinculacion/addparticipanteadministrativo.html", data)
                except Exception as ex:
                    pass

            elif action == 'deleteproyecto':
                try:
                    data['title'] = u'Eliminar Proyecto'
                    data['proyecto'] = ProyectosInvestigacion.objects.get(pk=request.GET['id'],tipo=1)
                    return render(request, "inv_vinculacion/deleteproyecto.html", data)
                except Exception as ex:
                    pass

            elif action == 'deletefecha':
                try:
                    data['title'] = u'Eliminar Convocatoria'
                    data['convocatoria'] = FechaProyectos.objects.get(pk=request.GET['id'])
                    return render(request, "inv_vinculacion/deletefecha.html", data)
                except Exception as ex:
                    pass
            elif action == 'deletecargobeneficiario':
                try:
                    data['title'] = u'Eliminar Cargo'
                    data['cargo'] = CargoBeneficiario.objects.get(pk=request.GET['id'])
                    return render(request, "inv_vinculacion/deletecargobeneficiario.html", data)
                except Exception as ex:
                    pass

            elif action == 'participantesproyectos':
                try:
                    participantes_estudiantes = []
                    panel = None
                    data['proyecto'] = proyectos = ProyectosInvestigacion.objects.get(pk=request.GET['id'], tipo=1)
                    data['tipoparticipante'] = ParticipantesTipo.objects.filter(tipo=proyectos.tipo)
                    participantesmatrices = ParticipantesMatrices.objects.filter(status=True, matrizevidencia_id=2,
                                                                                 proyecto=proyectos).order_by(
                        'inscripcion__persona__apellido1')
                    personal_docente = participantesmatrices.filter(
                        profesor__in=participantesmatrices.values_list('profesor_id', flat=True))
                    personal_administrativo = participantesmatrices.filter(
                        administrativo__in=participantesmatrices.values_list('administrativo_id', flat=True))
                    participantes_estudiantes = participantesmatrices.filter(
                        inscripcion__in=participantesmatrices.values_list('inscripcion_id', flat=True))
                    # participantes_estudiantes = participantesmatrices.filter(inscripcion__in=participantesmatrices.values_list('inscripcion_id', flat=True))
                    if 'panel' in request.GET:
                        panel = request.GET['panel']
                        if panel == '2':
                            data['personal_docente'] = personal_docente
                        elif panel == '3':
                            data['personal_administrativo'] = personal_administrativo
                    else:
                        # participantes_estudiantes = []
                        data['title'] = u'Participantes de Proyectos'
                        search = None
                        ids = None
                        inscripcionid = None
                        if 's' in request.GET:
                            search = request.GET['s']
                            if ' ' in search:
                                s = search.split(" ")
                                participantes_estudiantes = participantesmatrices.filter((
                                                                                                     Q(inscripcion__persona__apellido1__contains=
                                                                                                       s[0]) & Q(
                                                                                                 inscripcion__persona__apellido2__contains=
                                                                                                 s[1])))
                            else:
                                participantes_estudiantes = participantesmatrices.filter(
                                    Q(inscripcion__persona__nombres__contains=search) |
                                    Q(inscripcion__persona__apellido1__contains=search) |
                                    Q(inscripcion__persona__apellido2__contains=search) |
                                    Q(inscripcion__persona__cedula__contains=search))
                        else:
                            participantesmatrices = ParticipantesMatrices.objects.filter(status=True,
                                                                                         matrizevidencia_id=2,
                                                                                         proyecto=proyectos).order_by(
                                '-tipoparticipante', 'inscripcion__persona__apellido1')
                            participantes_estudiantes = participantesmatrices.filter(
                                inscripcion__in=participantesmatrices.values_list('inscripcion_id', flat=True))
                        paging = MiPaginador(participantes_estudiantes, 25)
                        p = 1
                        try:
                            paginasesion = 1
                            if 'paginador' in request.session:
                                paginasesion = int(request.session['paginador'])
                            if 'page' in request.GET:
                                p = int(request.GET['page'])
                            else:
                                p = paginasesion
                            try:
                                page = paging.page(p)
                            except:
                                p = 1
                            page = paging.page(p)
                        except:
                            page = paging.page(p)
                        request.session['paginador'] = p
                        data['paging'] = paging
                        data['page'] = page
                        data['rangospaging'] = paging.rangos_paginado(p)

                        data['search'] = search if search else ""
                        data['ids'] = ids if ids else ""
                        data['participantes'] = page.object_list
                    data['cantidad_estudiantes'] = participantes_estudiantes.count()
                    data['cantidad_docentes'] = personal_docente.count()
                    data['cantidad_administrativos'] = personal_administrativo.count()
                    data['panel'] = panel
                    data['proyectoid'] = proyectos.id
                    data['tieneperiodoinscripcion'] = PeriodoInscripcionVinculacion.objects.filter(status=True,proyecto_id=proyectos.id).exists()

                    return render(request, "inv_vinculacion/participantesproyectos.html", data)
                except Exception as ex:
                    pass
            elif action == 'revisarinformes':
                try:
                    data['id'] = id = int(request.GET['id'])
                    data['proyecto'] = matrizproyecto = ParticipantesMatrices.objects.get(pk=int(request.GET['id']))
                    data['part'] = participante = int(request.GET['participante'])
                    ESTADOS_PASOS = (
                        (1, u'APROBADO'),
                        (2, u'RECHAZADO')
                    )

                    ESTADOS_INFORMES = (
                        (1, u'APROBAR'),
                        (2, u'RECHAZAR')
                    )
                    data['informes_estudiantes'] = informes_estudiantes = InformesProyectoVinculacionEstudiante.objects.filter(status=True, proyecto=matrizproyecto).order_by('informedocente__nombre')
                    data['estados'] = ESTADOS_PASOS
                    data['estados_informes'] = ESTADOS_INFORMES

                    template = get_template('inv_vinculacion/modal/modal_revisarinformes.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, "mensaje": ex})

            elif action == 'participantesproyectostodos':
                try:
                    data['title'] = u' Estudiantes Participantes en todos los proyectos'
                    search = None
                    ids = None
                    inscripcionid = None

                    if 's' in request.GET:
                        search = request.GET['s']
                        if ' ' in search:
                            s = search.split(" ")
                            participantes = ParticipantesMatrices.objects.filter((Q(inscripcion__persona__apellido1__contains=s[0]) &
                                                                                  Q(inscripcion__persona__apellido2__contains=s[1])) ,
                                                                                 status=True, matrizevidencia_id=2 ).order_by('inscripcion__persona__apellido1')
                        else:
                            participantes = ParticipantesMatrices.objects.filter(Q(inscripcion__persona__nombres__contains=search) |
                                                                                 Q(inscripcion__persona__apellido1__contains=search) |
                                                                                 Q(inscripcion__persona__apellido2__contains=search) |
                                                                                 Q(inscripcion__persona__cedula__contains=search),
                                                                                 status=True, matrizevidencia_id=2).order_by('inscripcion__persona__apellido1')
                    else:
                        participantes = ParticipantesMatrices.objects.filter(status=True, matrizevidencia_id=2, profesor__isnull=True).order_by('-tipoparticipante', 'inscripcion__persona__apellido1')

                    paging = MiPaginador(participantes, 25)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['page'] = page
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['participantes'] = page.object_list
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    return render(request, "inv_vinculacion/participantesproyectostodos.html", data)
                except Exception as ex:
                    pass



            elif action == 'participantesproyectostodos_2':
                try:
                    data['title'] = u' Estudiantes Participantes en todos los proyectos'
                    search = None
                    ids = None
                    inscripcionid = None

                    if 's' in request.GET:
                        search = request.GET['s']
                        if ' ' in search:
                            s = search.split(" ")
                            participantes = InscripcionActividadConvalidacionPPV.objects.filter((Q(inscripcion__persona__apellido1__contains=s[0]) &
                                                                                  Q(inscripcion__persona__apellido2__contains=s[1])) ,
                                                                                 status=True, actividad__tipoactividad=2).order_by('inscripcion__persona__apellido1')
                        else:
                            participantes = InscripcionActividadConvalidacionPPV.objects.filter(Q(inscripcion__persona__nombres__contains=search) |
                                                                                 Q(inscripcion__persona__apellido1__contains=search) |
                                                                                 Q(inscripcion__persona__apellido2__contains=search) |
                                                                                 Q(inscripcion__persona__cedula__contains=search),
                                                                                 status=True, actividad__tipoactividad=2).order_by('inscripcion__persona__apellido1')
                    else:
                        participantes = InscripcionActividadConvalidacionPPV.objects.filter(status=True, actividad__tipoactividad=2).order_by('inscripcion__persona__apellido1')

                    paging = MiPaginador(participantes, 25)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['page'] = page
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['participantes'] = page.object_list
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    return render(request, "inv_vinculacion/participantesproyectostodos_2.html", data)
                except Exception as ex:
                    pass


            if action == 'cambiarinscripcion_2':
                try:
                    # id = None
                    # id = request.GET['id']
                    data['title'] = u'Cambiar Inscripción del Proyecto de Vinculación'
                    idest = int(request.GET['idest'])
                    if 'id' in request.GET:
                        idpy= int(request.GET['id'])
                    else:
                        return JsonResponse({"result": "bad", "mensaje": u"Debe contener una actividad"})
                    data['practica'] = practica = InscripcionActividadConvalidacionPPV.objects.get(inscripcion_id =idest,actividad_id=idpy, status = True)

                    participantes = ParticipantesMatrices.objects.get(inscripcion_id =idest, status=True, matrizevidencia_id=2, profesor__isnull=True)

                    data['participantes'] = participantes.id

                    form = CambioInscripcionVinculacionForm(initial={
                        'alumno': participantes.inscripcion.persona.nombre_completo_inverso(),
                        'carreraactual': participantes.inscripcion.carrera,
                        'activiades_cargadas': participantes.actividad,

                    })

                    # form.cargar_actividades(participantes)
                    form.cargar_otra_carrera(practica)
                    data['form'] = form
                    return render(request, "inv_vinculacion/editinscripcion.html", data)
                except Exception as ex:
                    pass

            # if action == 'cambiarinscripcion_2_2':
            #     try:
            #         data['title'] = u'Cambiar Inscripción del Proyecto de Vinculación'
            #         idest = int(request.GET['idest'])
            #         idpy= int(request.GET['id'])
            #         data['practica'] = practica = InscripcionActividadConvalidacionPPV.objects.get(pk=request.GET['id'])
            #
            #
            #         form = CambioInscripcionVinculacionForm(initial={
            #             'alumno': practica.inscripcion.persona.nombre_completo_inverso(),
            #             'carreraactual': practica.inscripcion.carrera,
            #         })
            #
            #         form.cargar_actividades(practica)
            #
            #         form.cargar_otra_carrera(practica)
            #
            #
            #         data['form'] = form
            #         return render(request, "inv_vinculacion/editinscripcion.html", data)
            #     except Exception as ex:
            #         pass


            elif action == 'participantesproyectosdocentes':
                try:
                    data['title'] = u' Docentes Participantes '
                    search = None
                    ids = None
                    inscripcionid = None
                    data['proyecto'] = proyectos = ProyectosInvestigacion.objects.get(pk=request.GET['id'],tipo=1)
                    data['tipoparticipante'] =  ParticipantesTipo.objects.filter(tipo=proyectos.tipo)
                    if 's' in request.GET:
                        search = request.GET['s']
                        if ' ' in search:
                            s = search.split(" ")
                            participantes = ParticipantesMatrices.objects.filter(Q(profesor__persona__apellido1__contains=s[0]) &
                                                                                  Q(profesor__persona__apellido2__contains=s[1]),
                                                                                 status=True, matrizevidencia_id=2,  proyecto=proyectos).order_by('profesor__persona__apellido1')
                        else:
                            participantes = ParticipantesMatrices.objects.filter( Q(profesor__persona__nombres__contains=search) |
                                                                                 Q(profesor__persona__apellido1__contains=search) |
                                                                                 Q(profesor__persona__apellido2__contains=search) |
                                                                                 Q(profesor__persona__cedula__contains=search) ,
                                                                                 status=True, matrizevidencia_id=2, proyecto=proyectos).order_by('profesor__persona__apellido1')
                    else:
                        participantes = proyectos.participantesmatrices_set.filter(status=True, matrizevidencia_id=2, inscripcion__isnull=True).order_by('-tipoparticipante', 'profesor__persona__apellido1')
                    paging = MiPaginador(participantes, 25)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['page'] = page
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['participantes'] = page.object_list
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['cambios'] = cambio=  ConfiguracionCambio.objects.filter(status=True, proyecto=proyectos, fecha_inicio__lte=datetime.now(),fecha_fin__gte=datetime.now()).exists()
                    if cambio:
                        data['configuracion'] = ConfiguracionCambio.objects.filter(status=True, proyecto=proyectos,fecha_inicio__lte=datetime.now(),fecha_fin__gte=datetime.now()).last()
                    return render(request, "inv_vinculacion/participantesproyectosdocentes.html", data)
                except Exception as ex:
                    pass

            elif action == 'carrerasproyectos':
                try:
                    data['title'] = u'Carreras de Proyecto'
                    data['proyecto'] = proyectos = ProyectosInvestigacion.objects.get(pk=request.GET['id'],tipo=1)
                    data['carreras'] = Carrera.objects.filter(status=True).exclude(coordinacion__id__in=[9, 7]).order_by('nombre')
                    data['carrerasproyecto'] = CarrerasProyecto.objects.filter(status=True, proyecto=proyectos)
                    return render(request, "inv_vinculacion/carrerasproyectos.html", data)
                except Exception as ex:
                    pass

            elif action == 'presupuestoproyectos':
                try:
                    data['title'] = u'Presupuesto de Proyecto'
                    data['proyecto'] = proyectos = ProyectosInvestigacion.objects.get(pk=request.GET['id'],tipo=1)
                    data['presupuestoproyecto'] = presupuesto = proyectos.presupuestosproyecto_set.filter(status=True) #PresupuestosProyecto.objects.filter(status=True, proyecto=proyectos)
                    data['totalplanificado'] = presupuesto.aggregate(totoplanificado=Sum('planificado'))['totoplanificado']
                    data['totalejecutado'] = presupuesto.aggregate(totejecutado=Sum('ejecutado'))['totejecutado']
                    data['formpresupuesto'] = PresupuestoProyectoForm
                    return render(request, "inv_vinculacion/presupuestoproyectos.html", data)
                except Exception as ex:
                    pass

            elif action == 'cargarsublineas':
                try:
                    # puede_realizar_accion(request, 'sga.puede_modificar_profesor_materia')
                    if 'linea_id' in request.GET:
                        lista = []
                        lineas= SubLineaInvestigacion.objects.filter(lineainvestigacion=int(request.GET['linea_id']))
                        for lis in lineas:
                            lista.append([lis.id, lis.nombre])
                        data = {"results": "ok", 'lista':lista}
                        return JsonResponse(data)
                except Exception as ex:
                    pass

            elif action == 'cargacanton':
                try:
                    # puede_realizar_accion(request, 'sga.puede_modificar_profesor_materia')
                    if 'zona_id' in request.GET:
                        lista = []
                        lineas= Canton.objects.filter(zona__id=int(request.GET['zona_id']))
                        for lis in lineas:
                            lista.append([lis.id, lis.nombre])
                        data = {"results": "ok", 'lista':lista}
                        return JsonResponse(data)
                except Exception as ex:
                    pass

            elif action == 'evidencias':
                try:
                    data['title'] = u'Formato Evidencias para los Proyectos'
                    data['evidencias'] = Evidencia.objects.filter(status=True, matrizevidencia_id=2)
                    return render(request, "inv_vinculacion/evidencia.html", data)
                except Exception as ex:
                    pass


            elif action == 'addevidenciaformato':
                try:
                    data['title'] = u'Adicionar Formato Evidencias para los Proyectos'
                    data['form'] = EvidenciaFormatoForm()
                    return render(request, "inv_vinculacion/addevidenciaformato.html", data)
                except Exception as ex:
                    pass

            elif action == 'editevidenciaformato':
                try:
                    data['title'] = u'Editar Formato Evidencias para los Proyectos'
                    data['evidencia'] = evidencia = Evidencia.objects.get(pk=request.GET['id'])
                    form = EvidenciaFormatoForm(initial={'nombre': evidencia.nombre})
                    data['form'] = form
                    return render(request, "inv_vinculacion/editevidenciaformato.html", data)
                except Exception as ex:
                    pass
            elif action == 'addinforme':
                try:
                    data['title'] = u'Nuevo Informe'
                    form = FechaInformeForm()
                    data['id'] = request.GET['id']
                    data['form'] = form

                    return render(request, "inv_vinculacion/addinforme.html", data)
                except Exception as ex:
                    pass

            elif action == 'addlinea':
                try:
                    data['programa'] = programa = ProgramasInvestigacion.objects.get(pk=request.GET['id'])
                    data['title'] = u'Linea Investigación ' + programa.nombre
                    data['lineaprogramasinvestigacions'] = LineaProgramasInvestigacion.objects.filter(status=True, programasinvestigacion=programa).order_by('-id')
                    data['areaprogramasinvestigacions'] = AreaProgramasInvestigacion.objects.filter(status=True, programasinvestigacion=programa).order_by('-id')
                    return render(request, "inv_vinculacion/addlinea.html", data)
                except Exception as ex:
                    pass

            # if action == 'addarea':
            #     try:
            #         data['programa'] = programa = ProgramasInvestigacion.objects.get(pk=request.GET['id'])
            #         data['title'] = u'Area de Conocimiento ' + programa.nombre
            #         data['areaprogramasinvestigacions'] = AreaProgramasInvestigacion.objects.filter(status=True, programasinvestigacion=programa).order_by('-id')
            #         return render(request, "inv_vinculacion/addarea.html", data)
            #     except Exception as ex:
            #         pass
            #
            elif action == 'addlineaprograma':
                try:
                    data['title'] = u'Adicionar lineas Investigación'
                    data['programa'] = programa = ProgramasInvestigacion.objects.get(pk=request.GET['id'])
                    data['form'] = ProgramasVinculacionLineaForm()
                    return render(request, "inv_vinculacion/addlineaprograma.html", data)
                except Exception as ex:
                    pass

            elif action == 'cambiarEstado':
                try:
                    programas = ProgramasInvestigacion.objects.get(pk=request.GET['id'], status=True)
                    if programas.estado:
                        programas.estado = False
                    else:
                        programas.estado = True
                    programas.save(request)
                    log(u'Cambio estado de programa: %s' % programas, request, "edit")
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

            elif action == 'addareaprograma':
                try:
                    data['title'] = u'Adicionar Area de Conocimiento'
                    data['programa'] = programa = ProgramasInvestigacion.objects.get(pk=request.GET['id'])
                    data['form'] = ProgramasVinculacionAreaForm()
                    return render(request, "inv_vinculacion/addareaprograma.html", data)
                except Exception as ex:
                    pass

            elif action == 'dellineaprograma':
                try:
                    data['title'] = u'Eliminar linea investigación en el programa'
                    data['lineaprogramasinvestigacion'] = LineaProgramasInvestigacion.objects.get(pk=request.GET['id'])
                    return render(request, "inv_vinculacion/dellineaprograma.html", data)
                except Exception as ex:
                    pass

            elif action == 'delareaprograma':
                try:
                    data['title'] = u'Eliminar area de conocimiento en el programa'
                    data['areaprogramasinvestigacion'] = AreaProgramasInvestigacion.objects.get(pk=request.GET['id'])
                    return render(request, "inv_vinculacion/delareaprograma.html", data)
                except Exception as ex:
                    pass

            elif action == 'addcarreraparticipante':
                try:
                    data['title'] = u'Adicionar'
                    data['form'] = CarreraParticipanteForm
                    data['idproyecto'] = request.GET['idproyecto']
                    return render(request, "inv_vinculacion/addcarreraparticipante.html", data)
                except Exception as ex:
                    pass

            elif action == 'editcarreraparticipante':
                try:
                    data['title'] = u'Adicionar'
                    data['id']= id= request.GET['id']
                    datos =  CarrerasParticipantes.objects.get(pk=id)
                    data['form'] = CarreraParticipanteForm(initial={
                        'carrera': datos.carrera,
                        # 'cupos': datos.cupos,
                    })
                    data['idproyecto'] = datos.proyecto.pk
                    return render(request, "inv_vinculacion/editcarreraparticipante.html", data)
                except Exception as ex:
                    pass

            elif action == 'deletecarreraparticipante':
                try:
                    data['title'] = u'Eliminar carrera participante'
                    data['participante'] = carrera = CarrerasParticipantes.objects.get(pk=request.GET['id'])
                    return render(request, "inv_vinculacion/deletecarreraparticipante.html", data)
                except Exception as ex:
                    pass

            elif action == 'addperfilprofesional':
                try:
                    data['title'] = u'Adicionar perfil profesional'
                    data['form'] = form = PerfilProfesionalForm()
                    carrera = CarrerasParticipantes.objects.get(pk=request.GET['idcarrera'])
                    form.filtro(carrera.carrera)
                    data['idproyecto'] = request.GET['idproyecto']
                    data['idcarrera'] = request.GET['idcarrera']
                    return render(request, "inv_vinculacion/addperfilprofesional.html", data)
                except Exception as ex:
                    pass

            elif action == 'editperfilprofesional':
                try:
                    data['title'] = u'Adicionar perfil profesional'
                    data['id'] = id = request.GET['id']
                    perfil = PerfilProfesional.objects.get(pk=id)
                    data['idproyecto'] = perfil.proyecto.pk
                    data['form'] = form = PerfilProfesionalForm(initial={
                        'perfil' : perfil.perfil,
                        'resultados_aprendizaje' : perfil.resultados_aprendizaje,
                        'asignatura': [i.id for i in perfil.asignatura.all()]
                    })
                    form.filtro(perfil.carreras_participantes.carrera)
                    return render(request, "inv_vinculacion/editperfilprofesional.html", data)
                except Exception as ex:
                    pass

            elif action == 'addbeneficiario':
                try:
                    data['title'] = u'Adicionar beneficiarios'
                    data['form'] = BeneficiariosForm
                    data['idproyecto'] = request.GET['idproyecto']
                    return render(request, "inv_vinculacion/addbeneficiario.html", data)
                except Exception as ex:
                    pass

            elif action == 'editbeneficiario':
                try:
                    data['title'] = u'Editar beneficiarios'
                    data['id'] = id = request.GET['id']
                    datos = Beneficiarios.objects.get(pk=id)
                    data['form'] = BeneficiariosForm(initial={
                        'nombre': datos.nombre,
                        'direccion': datos.direccion,
                        'correo': datos.correo,
                        'representante': datos.representante,
                        'cargo_repre': datos.cargo_repre,
                        'telefono': datos.telefono,
                        'coordenadas': datos.coordenadas,
                        'num_beneficiario_directo': datos.num_beneficiario_directo,
                        'num_beneficiario_indirecto': datos.num_beneficiario_indirecto,
                        'especifico': datos.especifico,
                        'caracteristica': datos.caracteristica,
                    })


                    data['idproyecto'] = datos.proyecto.id
                    return render(request, "inv_vinculacion/editbeneficiario.html", data)
                except Exception as ex:
                    pass

            elif action == 'deletebeneficiario':
                try:
                    data['title'] = u'Eliminar beneficiario'
                    data['participante'] = beneficiario = Beneficiarios.objects.get(pk=request.GET['id'])
                    return render(request, "inv_vinculacion/deletebeneficiario.html", data)
                except Exception as ex:
                    pass


            elif action == 'addinvolucrado':
                try:
                    data['title'] = u'Adicionar involucrados'
                    data['form'] = InvolucradoForm
                    data['idproyecto'] = request.GET['idproyecto']
                    return render(request, "inv_vinculacion/addinvolucrado.html", data)
                except Exception as ex:
                    pass

            elif action == 'editinvolucrado':
                try:
                    data['title'] = u'Editar involucrados'
                    data['id'] = id = request.GET['id']
                    datos = Involucrado.objects.get(pk=id)
                    data['form'] = InvolucradoForm(initial={
                        'nombre': datos.nombre,
                        'direccion': datos.direccion,
                        'correo': datos.correo,
                        'representante': datos.representante,
                        'cargo_repre': datos.cargo_repre,
                        'telefono': datos.telefono,
                        'coordenadas': datos.coordenadas,
                        'tipoActor': datos.tipoActor,
                    })


                    data['idproyecto'] = datos.proyecto.id
                    return render(request, "inv_vinculacion/editinvolucrado.html", data)
                except Exception as ex:
                    pass

            elif action == 'deleteinvolucrado':
                try:
                    data['title'] = u'Eliminar beneficiario'
                    data['participante'] = Involucrado.objects.get(pk=request.GET['id'])
                    return render(request, "inv_vinculacion/deleteinvolucrado.html", data)
                except Exception as ex:
                    pass


            elif action == 'editArbolProblema':
                try:
                    data['title'] = u'Editar árbol de problema'
                    data['id'] = id = request.GET['id']
                    datos = ArbolProblema.objects.get(id=id)
                    data['form'] = ArProblemaForm(initial={
                        'detalle': datos.detalle,
                    })
                    id = request.GET['id']
                    data['idproyecto'] = datos.proyecto.id
                    return render(request, "inv_vinculacion/editarbolproblema.html", data)
                except Exception as ex:
                    pass

            elif action == 'addSubCausa':
                try:
                    data['title'] = u'Añadir subcausa'
                    data['form'] = ArProblemaForm()
                    data['pry'] = request.GET['pry']
                    data['id'] = request.GET['id']
                    return render(request, "inv_vinculacion/addsubcausa.html", data)
                except Exception as ex:
                    pass

            elif action == 'addCausaEfecto':
                try:
                    data['title'] = u'Añadir causa efecto'
                    data['form'] = ArProb_CausaEfectoForm()
                    data['pry'] = request.GET['pry']
                    return render(request, "inv_vinculacion/addcausaefecto.html", data)
                except Exception as ex:
                    pass

            elif action == 'addDatoSecundario':
                    try:
                        data['title'] = u'Añadir dato secundario'
                        data['form'] = DatoSecundarioForm()
                        data['pry'] = request.GET['id']
                        return render(request, "inv_vinculacion/adddatosecundario.html", data)
                    except Exception as ex:
                        pass

            elif action == 'addcronograma':
                    try:
                        data['title'] = u'Añadir cronograma'
                        data['form'] = form = CronogramaForm()
                        data['aObj'] = request.GET['aObj']
                        aobj = ArbolObjetivo.objects.get(pk=request.GET['aObj'])
                        data['pry'] = aobj.proyecto.pk
                        docentes = ParticipantesMatrices.objects.values_list('profesor__pk',flat=True).filter(status=True, proyecto=aobj.proyecto)
                        form.fields["responsable"].queryset = Profesor.objects.filter(status=True,pk__in=docentes)
                        return render(request, "inv_vinculacion/addcronograma.html", data)
                    except Exception as ex:
                        pass

            elif action == 'editDatoSecundario':
                    try:
                        data['title'] = u'Editar dato secundario'
                        datos = DatoSecundario.objects.get(pk= request.GET['id'])
                        data['form'] = DatoSecundarioForm(
                            initial={
                                'descripcion':datos.descripcion,
                            }
                        )
                        data['id'] = request.GET['id']
                        data['pry'] = datos.proyecto.pk
                        return render(request, "inv_vinculacion/edidatosecundario.html", data)
                    except Exception as ex:
                        pass

            elif action == 'editLineaBase':
                    try:
                        data['title'] = u'Editar dato primario'
                        data['id'] = id =  request.GET['id']
                        datos = MatrizLineaBase.objects.get(pk= id)
                        data['form'] = LineaBaseForm(
                            initial={
                                'item': datos.arbolProblema.detalle,
                                'descripcion': datos.descripcion,
                                'metodo': datos.metodo,
                                'fuente': datos.fuente,
                                'linea_base': datos.datos_linea_base,
                            }
                        )

                        data['pry'] = datos.proyecto.pk
                        return render(request, "inv_vinculacion/editlineabase.html", data)
                    except Exception as ex:
                        pass

            elif action == 'editArbolObjetivo':
                    try:
                        data['title'] = u'Editar árbol de objetivo'
                        data['id'] = id =  request.GET['id']
                        datos = ArbolObjetivo.objects.get(id=id)
                        data['form'] = ArObjetivoForm(
                            initial={
                                 'detalle': datos.detalle,
                            }
                        )

                        data['pry'] = datos.proyecto.pk
                        return render(request, "inv_vinculacion/editarbolobjetivo.html", data)
                    except Exception as ex:
                        pass

            elif action == 'editMarcoLogico':
                    try:
                        data['title'] = u'Editar marco lógico'
                        data['id'] = id =  request.GET['id']
                        datos = MarcoLogico.objects.get(id=id)
                        if request.GET['bool'] == 'True':
                            data['bool_marLogi'] = True
                        data['form']= form = MarcoLogicoForm2(
                            initial={
                                'resumen_narrativo': datos.resumen_narrativo,
                                'dato': MatrizLineaBase.objects.get(arbolProblema=datos.arbolObjetivo.arbolProblema).datos_linea_base  if MatrizLineaBase.objects.get(arbolProblema=datos.arbolObjetivo.arbolProblema).datos_linea_base else "No existe línea base",
                                'indicador': datos.indicador,
                                'fuente': datos.fuente,
                                'supuestos': datos.supuestos,
                                'cumplimiento': datos.cumplimiento,
                            }
                        )
                        if datos.arbolObjetivo.parentID:
                            form.deshabilitar_campo('cumplimiento')
                        data['pry'] = datos.proyecto.pk
                        return render(request, "inv_vinculacion/editmarcologico2.html", data)
                    except Exception as ex:
                        pass

            elif action == 'deleteArbolProblema':
                try:
                    data['title'] = u'Eliminar detalle arbol de problema'
                    data['detalle'] = ArbolProblema.objects.get(id=request.GET['id'])
                    return render(request, "inv_vinculacion/deletearbolproblema.html", data)
                except Exception as ex:
                    pass

            elif action == 'deleteDatoSecundario':
                try:
                    data['title'] = u'Eliminar dato secundario'
                    data['detalle'] = DatoSecundario.objects.get(id=request.GET['id'])
                    return render(request, "inv_vinculacion/deletedatosecundario.html", data)
                except Exception as ex:
                    pass

            elif action == 'editcronograma':
                    try:
                        data['title'] = u'Editar cronograma'
                        data['id'] = id = request.GET['id']
                        datos = Cronograma.objects.get(pk=id)
                        data['form'] = form = CronogramaForm(
                            initial={
                                'responsable': [i.id for i in datos.responsable.all()],
                                'descripcion': datos.descripcion,
                                'fecha_inicio': datos.fecha_inicio,
                                'fecha_fin': datos.fecha_fin,
                            }
                        )
                        data['pry'] = datos.proyecto.pk
                        docentes = ParticipantesMatrices.objects.values_list('profesor__pk',flat=True).filter(status=True, proyecto=datos.proyecto)
                        form.fields["responsable"].queryset = Profesor.objects.filter(status=True,pk__in=docentes)
                        return render(request, "inv_vinculacion/editcronograma.html", data)
                    except Exception as ex:
                        pass

            elif action == 'deletecronograma':
                try:
                    data['title'] = u'Eliminar cronograma'
                    data['detalle'] = Cronograma.objects.get(pk=request.GET['id'])
                    return render(request, "inv_vinculacion/deletecronograma.html", data)
                except Exception as ex:
                    pass

            elif action == 'addpresupuesto':
                    try:
                        data['title'] = u'Añadir presupuesto'
                        data['form'] = form = PresupuestoForm()
                        data['pry'] = ArbolObjetivo.objects.get(pk=request.GET['aObj']).proyecto.pk
                        data['aObj'] = request.GET['aObj']
                        return render(request, "inv_vinculacion/addpresupuesto.html", data)
                    except Exception as ex:
                        pass

            elif action == 'editpresupuesto':
                    try:
                        data['title'] = u'Editar presupuesto'
                        data['id']=id= request.GET['id']
                        datos = Presupuesto.objects.get(pk=id)

                        data['form'] = form = PresupuestoForm(initial={
                            'cantidad': datos.cantidad,
                            'suministro': datos.suministro,
                            'costo_unitario': datos.costo_unitario,
                            'subtotal': datos.subtotal,
                            'iva': datos.iva,
                            'total': datos.total,
                            'especificaciones': datos.especificaciones,
                            'aplica_iva': datos.suministro.aplicaIva if datos.suministro else False
                        })
                        data['pry'] = datos.proyecto.pk
                        return render(request, "inv_vinculacion/editpresupuesto.html", data)
                    except Exception as ex:
                        pass

            elif action == 'deletepresupuesto':
                try:
                    data['title'] = u'Eliminar Presupuesto'
                    data['detalle'] = Presupuesto.objects.get(pk=request.GET['id'])
                    return render(request, "inv_vinculacion/deletepresupuesto.html", data)
                except Exception as ex:
                    pass

            elif action == 'addredaccion':
                    try:
                        data['title'] = u'Añadir redacción'
                        data['form'] = form = RedaccionForm()
                        data['pry'] = request.GET['idproyecto']
                        return render(request, "inv_vinculacion/addredaccion.html", data)
                    except Exception as ex:
                        pass

            elif action == 'editredaccion':
                    try:
                        data['title'] = u'Editar redacción'
                        data['id'] = id = request.GET['id']
                        datos = Problema.objects.get(pk=id)
                        data['form'] = form = RedaccionForm(
                            initial={
                                'proyecto': datos.proyecto,
                                'antecedentes': datos.antecedentes,
                                'descripcion': datos.descripcion,
                                'deteccion_necesidades': datos.deteccion_necesidades,
                                'justificacion': datos.justificacion,
                                'Pertinencia': datos.Pertinencia,
                                'metodologia': datos.metodologia,
                                'seguimiento': datos.seguimiento,
                                'evaluacion': datos.evaluacion,
                                'producto': datos.producto,
                                'Bibliografia': datos.Bibliografia,
                                'objetivo_general': datos.objetivo_general,
                                'objetivos_especificos': datos.objetivos_especificos,
                                'convenio': datos.convenio,
                            }
                        )
                        data['pry'] = datos.proyecto.pk
                        return render(request, "inv_vinculacion/editredaccion.html", data)
                    except Exception as ex:
                        pass


            elif action == 'addanexo':
                    try:
                        data['title'] = u'Añadir anexos'
                        data['form'] = form = AnexosForm()
                        data['pry'] = request.GET['idproyecto']
                        return render(request, "inv_vinculacion/addanexo.html", data)
                    except Exception as ex:
                        pass

            elif action == 'editanexo':
                    try:
                        data['title'] = u'Editar anexos'
                        data['id'] = id= request.GET['id']
                        datos = Anexos.objects.get(pk=id)
                        data['form'] = form = AnexosForm(
                            initial={
                                'titulo': datos.titulo,
                            }
                        )
                        data['pry'] = datos.proyecto.pk
                        return render(request, "inv_vinculacion/editanexo.html", data)
                    except Exception as ex:
                        pass

            elif action == 'deleteanexo':
                try:
                    data['title'] = u'Eliminar anexo'
                    data['detalle'] = Anexos.objects.get(pk=request.GET['id'])
                    return render(request, "inv_vinculacion/deleteanexo.html", data)
                except Exception as ex:
                    pass

            elif action == 'addsuministro':
                    try:
                        data['title'] = u'Añadir suministro'
                        data['form'] = form = SuministroForm()
                        return render(request, "inv_vinculacion/addsuministro.html", data)
                    except Exception as ex:
                        pass

            elif action == 'editsuministro':
                    try:
                        data['title'] = u'Editar suministro'
                        data['id'] = id = request.GET['id']
                        datos = Suministro.objects.get(pk=id)
                        data['form'] = form = SuministroForm(
                            initial={
                                'rubro': datos.rubro,
                                'especificacion': datos.especificacion,
                                'costo_unitario': datos.costo_unitario,
                                'aplica_iva': datos.aplicaIva,
                                'activo': datos.activo,
                            }
                        )
                        return render(request, "inv_vinculacion/editsuministro.html", data)
                    except Exception as ex:
                        pass

            elif action == 'deletesuministro':
                try:
                    data['title'] = u'Eliminar suministro'
                    data['detalle'] = Suministro.objects.get(pk=request.GET['id'])
                    return render(request, "inv_vinculacion/deletesuministro.html", data)
                except Exception as ex:
                    pass

            elif action == 'costo':
                try:
                    suministro = Suministro.objects.get(status=True, pk=request.GET['id'])
                    costo = suministro.costo_unitario
                    coninva = suministro.aplicaIva
                    especificacion = suministro.especificacion
                    return JsonResponse({"result": "ok","costo":costo, 'aplicaIva': coninva,'especificacion': especificacion})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": u"Error al consultar los datos."})

            elif action == 'configurarinforme_adm':
              try:
                  data['title'] = u'Configurar Informe'
                  data['proyecto'] = proyecto= ProyectosInvestigacion.objects.get(pk=request.GET['id'])
                  data['configuraciones'] = configuracion_lider = ConfiguracionInformeVinculacion.objects.filter(status=True, proyecto= proyecto, aprobacion__in=[2,3,4], informecondensado=True).order_by('-fecha_inicio')
                  # data['lider'] = Persona.objects.get(pk=int(''.join(map(str, configuracion_lider.values_list('profesor__persona_id')[0]))))
                  lider = ParticipantesMatrices.objects.filter(status=True, proyecto=proyecto, tipoparticipante__pk=1).values_list('profesor_id', flat=True)
                  data['informePromotor'] = ConfiguracionInformeVinculacion.objects.filter(status=True, proyecto=proyecto).exclude(profesor_id__in=lider).order_by('-fecha_inicio')
                  data['tecnicoasociado'] = proyecto.get_tecnicoasociado()
                  return render(request, "inv_vinculacion/configuracion_adm.html", data)
              except Exception as ex:
                  pass

            elif action == 'generar_adm':
                try:
                    with transaction.atomic():
                        data['configuracion']= conf = ConfiguracionInformeVinculacion.objects.get(id=request.GET['id'])
                        data['promotor'] = Persona.objects.get(pk=conf.profesor.persona.id)
                        data['acciones'] = DetalleInformeVinculacion.objects.filter(status=True, configuracion__id=(request.GET['id']))
                        data['componentes'] = MarcoLogico.objects.filter(status=True, proyecto=conf.proyecto, arbolObjetivo__tipo=2, arbolObjetivo__parentID__isnull=True)
                        data['proyecto'] = conf.proyecto
                        proyecto = ProyectosInvestigacion.objects.get(pk=conf.proyecto.pk)
                        data['fines'] = DetalleInformeVinculacion.objects.filter(status=True, configuracion=conf, proyecto=proyecto, actividad__arbolObjetivo__tipo=3)
                        data['propositos'] = DetalleInformeVinculacion.objects.filter(status=True, configuracion=conf, proyecto=proyecto, actividad__arbolObjetivo__tipo=1)
                        data['aPro_marcoLogico_acciones'] = MarcoLogicoReporte.objects.filter(status=True, configuracion=conf, proyecto=proyecto, arbolObjetivo__tipo=2, arbolObjetivo__parentID__isnull=False)
                        data['aPro_marcoLogico_componentes'] = MarcoLogicoReporte.objects.filter(status=True, configuracion=conf, proyecto=proyecto,arbolObjetivo__tipo=2, arbolObjetivo__parentID__isnull=True)
                        return render(request, 'inv_vinculacion/generarinforme_adm.html', data)
                except Exception as ex:
                    pass

            elif action == 'aprobarconfiguracion':
                try:
                    data['title'] = u'Aprobacion de Configuraciones '
                    data['form'] = AprobacionInformeForm()
                    data['id'] = int(request.GET['id'])
                    data['proyecto'] = request.GET['idproyecto']
                    template = get_template("inv_vinculacion/aprobacion_conf.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content})
                except Exception as ex:
                    pass

            elif action == 'subirInforme':
                try:
                    data['title'] = u'Subir informe '
                    data['form2'] = SubirInformeForm()
                    data['id'] = int(request.GET['id'])
                    data['pry'] = int(request.GET['pry'])
                    template = get_template("inv_vinculacion/subirinforme.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content})
                except Exception as ex:
                    pass

            elif action == 'generarinforme_adm_pdf':
                try:

                    data['configuracion'] =  conf = ConfiguracionInformeVinculacion.objects.get(id=request.GET['id'])
                    profesor = Profesor.objects.get(persona__usuario=conf.usuario_creacion)
                    data['lider'] = ParticipantesMatrices.objects.get(status=True, proyecto=conf.proyecto,tipoparticipante__pk=1, activo = True)

                    detalle = DetalleInformeVinculacion.objects.filter(status=True, configuracion__id=request.GET['id'], editado=False).exclude(actividad__arbolObjetivo__tipo=3)
                    detalleinfo = DetalleInformeVinculacion.objects.filter(status=True, configuracion__id=request.GET['id'])
                    data['acciones'] = DetalleInformeVinculacion.objects.filter(status=True,usuario_creacion=conf.usuario_creacion,configuracion__id=(request.GET['id']))
                    data['componentes'] = MarcoLogico.objects.filter(status=True, proyecto=conf.proyecto,arbolObjetivo__tipo=2,arbolObjetivo__parentID__isnull=True).order_by('arbolObjetivo__orden')
                    data['proyecto'] = conf.proyecto
                    proyecto = ProyectosInvestigacion.objects.get(pk=conf.proyecto.pk)
                    data['fines'] = DetalleInformeVinculacion.objects.filter(status=True, configuracion=conf, proyecto=proyecto, usuario_creacion=conf.usuario_creacion,actividad__arbolObjetivo__tipo=3)
                    data['propositos'] = DetalleInformeVinculacion.objects.filter(status=True, configuracion=conf, proyecto=proyecto, usuario_creacion=conf.usuario_creacion,actividad__arbolObjetivo__tipo=1)
                    data['aPro_marcoLogico_acciones'] = MarcoLogicoReporte.objects.filter(status=True, configuracion=conf, proyecto=proyecto,arbolObjetivo__tipo=2,arbolObjetivo__parentID__isnull=False)
                    data['aPro_marcoLogico_componentes'] = MarcoLogicoReporte.objects.filter(status=True, configuracion=conf,proyecto=proyecto, arbolObjetivo__tipo=2,arbolObjetivo__parentID__isnull=True).order_by('arbolObjetivo__orden')
                    data['eslider'] = eslider = ParticipantesMatrices.objects.filter(status=True, proyecto=proyecto,tipoparticipante__pk=1,profesor=conf.profesor, activo = True).exists()

                    if eslider:
                        lider = ParticipantesMatrices.objects.get(status=True, proyecto=proyecto, tipoparticipante__pk=1, activo = True)
                        data['docente'] = lider.profesor.persona
                    else:
                        data['docente'] = profesor.persona

                    if responsablevinculacion:
                        data['responsablevinculacion'] = responsablevinculacion

                    if conf.personaaprueba == proyecto.lider():
                        data['revisor'] = 'LÍDER DE PROYECTO'
                    else:
                        data['revisor'] = 'TÉCNICO DOCENTE DE VINCULACIÓN'

                    return conviert_html_to_pdf(
                        'inv_vinculacion/generarinforme_adm_pdf.html',
                        {
                            'pagesize': 'A4',
                            'data': data,
                        }
                    )
                except Exception as ex:
                    pass
            elif action == 'listaconfigurarcambio':
                try:
                    data['title'] = 'Configuraciones de Cambio'
                    data ['configuraciones'] = ConfiguracionCambio.objects.filter(status=True)
                    return render(request, "inv_vinculacion/listadoconfigurarcambio.html", data)
                except Exception as ex:
                    pass
            elif action == 'addcambioconfiguracion':
                try:
                    data['title'] = u'Nuevo Cambio de Configuración'
                    form = ConfiguracionCambioForm()
                    data['form'] = form
                    template = get_template("inv_vinculacion/addconfiguracioncambio.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content})
                except Exception as ex:
                    pass
            elif action == 'editcambioconfiguracion':
                try:
                    data['title'] = u'Editar Cambio de Configuración'
                    data['id'] = id = request.GET['id']
                    cambio = ConfiguracionCambio.objects.get(id=id)
                    form = ConfiguracionCambioForm(initial={'proyecto': cambio.proyecto,
                                                           'fecha_inicio': cambio.fecha_inicio,
                                                           'fecha_fin': cambio.fecha_fin,
                                                            'tipo': cambio.tipo})
                    data['form'] = form
                    template = get_template("inv_vinculacion/editcambioconfiguracion.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content})
                except Exception as ex:
                    pass

            elif action == 'editfechagestion':
                try:
                    data['title'] = u'Editar Fecha de Gestión'
                    data['id'] = id = request.GET['id']
                    docente = ParticipantesMatrices.objects.get(id=id)
                    form = FechaGestionForm(initial={'fecha_inicio':docente.fecha_inicio,
                                                     'fecha_fin':docente.fecha_fin,
                                                     'activo': docente.activo
                                                     })
                    data['form'] = form
                    data['id_proy'] = docente.proyecto.pk
                    template = get_template("inv_vinculacion/editfechagestion.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content})
                except Exception as ex:
                    pass

            elif action == 'cambioDocente':
                try:
                    data['title'] = u'Cambio de docnete participante'
                    data['id'] = id = request.GET['id']
                    data['form'] = form = CambioDocenteVinculacionForm()
                    proyecto = ProyectosInvestigacion.objects.get(pk=id)
                    docentes = ParticipantesMatrices.objects.values_list('profesor__pk', flat=True).filter(status=True,proyecto=proyecto, activo=True)
                    form.fields["docente"].queryset = Profesor.objects.filter(status=True, pk__in=docentes)
                    template = get_template("inv_vinculacion/cambioDocente.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content})
                except Exception as ex:
                    pass

            elif action == 'addperiodoinscripcion':
                try:
                    data['title'] = u'Adicionar periodo inscripción'
                    data['form'] = PeriodoInscripcionFrom(initial={'periodo': periodo})
                    data['id'] = request.GET['id']
                    template = get_template("inv_vinculacion/addperiodoinscripcion.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content})
                except Exception as ex:
                    pass

            elif action == 'editperiodoinscripcion':
                try:
                    data['title'] = u'Editar periodo inscripción'
                    periodo = PeriodoInscripcionVinculacion.objects.get(pk= request.GET['id'])
                    data['form'] = PeriodoInscripcionFrom(initial={
                        'periodo': periodo.periodo,
                        'observacion': periodo.observacion,
                        'fechainicio': periodo.fechainicio,
                        'fechafin': periodo.fechafin,
                    })
                    data['id'] = request.GET['id']
                    data['proyecto'] = periodo.proyecto.pk
                    template = get_template("inv_vinculacion/editperiodoinscripcion.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content})
                except Exception as ex:
                    pass

            elif action == 'verdetalle':
                try:
                    data['title'] = u'Detalle carrera'

                    data['periodoinscrip'] = periodovinculacion = PeriodoInscripcionVinculacion.objects.get(pk=request.GET['id'])
                    data['carreras'] = CarreraInscripcionVinculacion.objects.select_related().filter(status=True, periodo = periodovinculacion)
                    template = get_template("inv_vinculacion/detalleinscripcion.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content})
                except Exception as ex:
                    pass

            elif action == 'inscripcion':
                try:
                    data['title'] = u'Inscripcion de estudiantes'
                    search = None
                    id = request.GET['id']
                    if 's' in request.GET:
                        search = request.GET['s']
                        periodoinscripcion = PeriodoInscripcionVinculacion.objects.select_related().filter(status=True, proyecto__pk = id, observacion__contains=search)
                    elif 'id' in request.GET:
                        periodoinscripcion = PeriodoInscripcionVinculacion.objects.select_related().filter(status=True, proyecto__pk=id)
                    paging = MiPaginador(periodoinscripcion, 25)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['page'] = page
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['periodos'] = page.object_list
                    data['proyecto'] = id
                    data['search'] = search if search else ""

                    return render(request, "inv_vinculacion/periodoInscripcion.html", data)
                except Exception as ex:
                    pass

            elif action == 'deletesolicitudproyecto':
                try:
                    data['title'] = u'Eliminar Solicitud'
                    data['participante'] = participante = ProyectoVinculacionInscripcion.objects.get(pk=request.GET['id'])
                    return render(request, "proyectovinculaciondocente/deletesolicitudproyecto.html", data)
                except Exception as ex:
                    pass


            elif action == 'verinscritos':
                try:
                    data['title'] = u'Solicitudes de Proyectos'
                    search = None
                    ids = None
                    inscripcionid = None
                    periodo = PeriodoInscripcionVinculacion.objects.get(pk=request.GET['id'])
                    data['proyecto'] = proyectos = ProyectosInvestigacion.objects.get(pk= periodo.proyecto.pk)
                    data['periodo'] = periodo
                    carreras = []
                    carreras = CarreraInscripcionVinculacion.objects.filter(status=True,periodo=periodo).values_list('carrera__carrera__id', 'carrera__carrera__nombre', flat=False)
                    data['carreras'] = carreras
                    participantes = ProyectoVinculacionInscripcion.objects.filter(status=True,periodo=periodo).order_by('inscripcion__persona__apellido1')
                    data['total'] = participantes.count()
                    car=0
                    est=0
                    if 'car' in request.GET:
                        if not request.GET['car'] == "0":
                            participantes = participantes.filter(inscripcion__carrera=request.GET['car'])
                            car = int(request.GET['car'])

                    if 'est' in request.GET:
                        if not request.GET['est'] == "0":
                            participantes = participantes.filter(estado=request.GET['est'])
                            est = int(request.GET["est"])

                    if 's' in request.GET:
                        search = request.GET['s']
                        if ' ' in search:
                            s = search.split(" ")
                            participantes = participantes.filter((Q(inscripcion__persona__apellido1__contains=s[0]) &
                                                                                            Q(inscripcion__persona__apellido2__contains=s[1])),
                                                                                            status=True, periodo=periodo).order_by('inscripcion__persona__apellido1')
                        else:
                            participantes = participantes.filter(Q(inscripcion__persona__nombres__contains=search) |
                                                                                        Q(inscripcion__persona__apellido1__contains=search) |
                                                                                        Q(inscripcion__persona__apellido2__contains=search) |
                                                                                        Q(inscripcion__persona__cedula__contains=search),
                                                                                        status=True, periodo=periodo).order_by('inscripcion__persona__apellido1')


                    paging = MiPaginador(participantes, 25)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['page'] = page
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['participantes'] = page.object_list
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['car'] = car
                    data['est'] = est
                    data['busqueda'] = participantes.count()
                    data['aprobadas'] = participantes.filter(estado=2).count()
                    data['pendientes'] = participantes.filter(estado=1).count()
                    return render(request, "inv_vinculacion/verinscritos.html", data)
                except Exception as ex:
                    pass

            elif action == 'definirregistrohoras':
                try:
                    titulo = 'Definir fechas de registro de horas'
                    data['form'] = FechasRegistroHorasForm
                    data['idproyecto'] = request.GET['id']
                    template = get_template("inv_vinculacion/modal/modal_definirregistrohoras.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", "titulo": titulo, 'data': json_content})
                except Exception as ex:
                    pass

            elif action == 'habilitarcargainforme':
                try:
                    data['title'] = u'Informes Para Proyecto de Vinculación'
                    data['id'] = id = request.GET['id']
                    data['proyecto'] = proyecto = ProyectosInvestigacion.objects.get(pk=request.GET['id'], tipo = 1)
                    data['fechaactual'] =  datetime.now().date()
                    data['informes_habilitados'] = InformesProyectoVinculacionDocente.objects.filter(status=True, proyecto=proyecto).order_by('fecha_creacion')
                    return render(request, 'inv_vinculacion/listadoinformeshabilitados.html', data)
                except Exception as ex:
                    pass

            elif action == 'addhabilitacioninforme':
                try:
                    data['id'] = id = request.GET['id']
                    data['proyecto'] = ProyectosInvestigacion.objects.get(pk=request.GET['id'], tipo = 1)
                    data['formulario'] = InformeProyectoVinculacionForm()
                    template = get_template("inv_vinculacion/modal/modal_habilitarinforme.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'editarhabilitacioninforme':
                try:
                    data['id'] = request.GET['id']

                    data['informe'] = informe = InformesProyectoVinculacionDocente.objects.get(pk=int(request.GET['id']))
                    data['formulario'] = InformeProyectoVinculacionForm(initial=model_to_dict(informe))
                    template = get_template("inv_vinculacion/modal/modal_habilitarinforme.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'eliminarhabilitacioninforme':
                try:
                    data['title'] = u'Eliminar informe habilitado'
                    data['informe'] = InformesProyectoVinculacionDocente.objects.get(pk=int(request.GET['id']))
                    return render(request, "inv_vinculacion/modal/modal_eliminarinformehabilitado.html", data)
                except Exception as ex:
                    pass

            elif action == 'habilitarregistrohoras':
                try:
                    data['title'] = u'Registro de Horas para Participantes de Proyecto de Vinculación'
                    data['id'] = id = request.GET['id']
                    data['proyecto'] = proyecto = ProyectosInvestigacion.objects.get(pk=request.GET['id'])
                    data['habilitaciones_horas_registradas'] = HabilitacionesHorasParticipantesVinculacion.objects.filter(status=True, proyecto=proyecto).order_by('-fecha_creacion')
                    return render(request, 'inv_vinculacion/listadohorashabilitadas.html', data)
                except Exception as ex:
                    pass

            elif action == 'addhabilitacionhoras':
                try:
                    data['id'] = id = request.GET['id']
                    data['proyecto'] = ProyectosInvestigacion.objects.get(pk=id)
                    data['convocatorias_periodos'] = PeriodoInscripcionVinculacion.objects.filter(status=True, proyecto_id=id, aprobado=True)
                    data['formulario'] = HabilitacionRegistroHorasVinculacionForm()
                    template = get_template("inv_vinculacion/modal/modal_habilitarhoras.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'editarhabilitacionhoras':
                try:
                    data['habilitacion_horas'] = habilitacion = HabilitacionesHorasParticipantesVinculacion.objects.get(pk=int(request.GET['id']))
                    data['convocatorias_periodos'] = PeriodoInscripcionVinculacion.objects.filter(status=True, proyecto_id=habilitacion.proyecto.id, aprobado=True)
                    data['formulario'] = HabilitacionRegistroHorasVinculacionForm(initial=model_to_dict(habilitacion))
                    template = get_template("inv_vinculacion/modal/modal_habilitarhoras.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'eliminarhabilitacionhoras':
                try:
                    data['title'] = u'Eliminar horas habilitadas'
                    data['habilitacion_horas'] = HabilitacionesHorasParticipantesVinculacion.objects.get(pk=int(request.GET['id']))
                    return render(request, "inv_vinculacion/modal/modal_eliminarhorashabilitadas.html", data)
                except Exception as ex:
                    pass

            elif action == 'addinscritomanual':
                try:
                    data['id'] = id = request.GET['id']
                    data['filtro'] = filtro = PeriodoInscripcionVinculacion.objects.get(pk=id)
                    form = InscripcionManualProyectoVinculacionForm()
                    form.fields['persona'].queryset = Persona.objects.none()
                    data['form2'] = form
                    template = get_template("inv_vinculacion/modal/formmodal.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'buscarinscritos':
                try:
                    id = request.GET['id']
                    q = request.GET['q'].upper().strip()
                    s = q.split(" ")
                    carrera = ""
                    actividad = PeriodoInscripcionVinculacion.objects.get(pk=id)
                    carreraid = CarrerasParticipantes.objects.values_list('carrera_id', flat=True).filter(status=True, proyecto=actividad.proyecto)

                    querybase = Inscripcion.objects.filter(status=True, activo=True, carrera__in=carreraid).order_by('persona__apellido1')

                    if len(s) == 1:
                        querybase = querybase.filter((Q(persona__nombres__icontains=q) | Q(persona__apellido1__icontains=q) | Q(persona__cedula__icontains=q) |  Q(persona__apellido2__icontains=q) | Q(persona__cedula__contains=q)), Q(status=True, perfilusuario__status=True,
                                                                 perfilusuario__visible=True)).distinct()[:15]
                    elif len(s) == 2:
                        querybase = querybase.filter((Q(persona__apellido1__contains=s[0]) & Q(persona__apellido2__contains=s[1])) |
                                                       (Q(persona__nombres__icontains=s[0]) & Q(persona__nombres__icontains=s[1])) |
                                                       (Q(persona__nombres__icontains=s[0]) & Q(persona__apellido1__contains=s[1]))).filter(status=True, perfilusuario__status=True,
                                                                 perfilusuario__visible=True).distinct()[:15]
                    else:
                        querybase = querybase.filter((Q(persona__nombres__contains=s[0]) & Q(persona__apellido1__contains=s[1]) & Q(persona__apellido2__contains=s[2])) |
                                                       (Q(persona__nombres__contains=s[0]) & Q(persona__nombres__contains=s[1]) & Q(persona__apellido1__contains=s[2]))).filter(status=True, perfilusuario__status=True,
                                                                 perfilusuario__visible=True).distinct()[:15]
                    data = {"result": "ok", "results": [{"id": x.id, "name": "{} - {} | {} | NIVEL MALLA {}".format(x.persona.cedula, x.persona.nombre_completo(), x.carrera.nombre, x.mi_nivel().nivel.orden)} for x in querybase]}
                    return JsonResponse(data)
                except Exception as ex:
                    pass

            elif action == 'editcarrera':
                try:
                    data['id'] = id = request.GET['id']
                    data['participante'] = participante = ParticipantesMatrices.objects.get(pk=int(id))

                    form = CambioCarreraVinculacionForm(initial={
                        'alumno': participante.inscripcion.persona.nombre_completo_inverso(),
                        'proyecto': participante.proyecto.nombre,
                        'horas': participante.horas,
                        'carreraactual': participante.inscripcion.carrera.nombre,
                    })

                    form.cargar_otra_carrera(participante)
                    data['form'] = form
                    data['action'] = action
                    template = get_template('inv_vinculacion/cambiarcarrera.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, "mensaje": ex})

            elif action == 'registrarhoras':
                try:
                    data['participante'] = participante = ParticipantesMatrices.objects.get(pk=int(request.GET['id']))
                    data['action'] = action
                    data['id'] = request.GET['id']

                    form = RegistrarHorasVinculacionForm(initial={
                        'alumno': participante.inscripcion.persona.nombre_completo_inverso(),
                        'proyecto': participante.proyecto.nombre,
                        'horas': participante.horas,
                        'carrera': participante.inscripcion.carrera.nombre,
                        'estado': participante.estado,
                        'fechainicio': participante.registrohorasdesde if participante.registrohorasdesde else datetime.now().date(),
                        'fechafinalizacion': participante.registrohorashasta if participante.registrohorashasta else datetime.now().date(),
                    })
                    # if not participante.cantidad_informes_aprobados() == participante.cantidad_informes_solicitados():
                    #     form.bloqueo('horas')
                    data['form'] = form
                    template = get_template('inv_vinculacion/registrarhorasvinculacion.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, "mensaje": ex})

            elif action == 'subirarchivoinscripcion':
                try:
                    data['id'] = id = int(encrypt(request.GET['id']))
                    data['filtro'] = filtro = PeriodoInscripcionVinculacion.objects.get(pk=id)
                    form = MasivoInscripcionVinculacionForm()
                    data['form'] = form
                    data['action'] = 'subirarchivoinscripcion'
                    template = get_template("inv_vinculacion/modal/formMigrarVinculacionMasivo.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": "bad", "message": u"Error al obtener los datos."})

            elif action == 'adddocenteexterno':
                try:
                    data['title'] = u'Agregar Docente Externo'
                    proyecto = ProyectosInvestigacion.objects.get(pk=request.GET['idp'])
                    form = DocenteExternoForm(initial = {'funcionproyecto' : ParticipantesTipo.objects.get(pk=5).nombre})
                    data['proyecto'] = proyecto
                    data['form'] = form
                    return render(request, "inv_vinculacion/adddocenteexterno.html", data)
                except Exception as ex:
                    pass

            elif action == 'verificarfirmas':
                try:
                    id = int(request.GET['id'])
                    informe_estudiante = InformesProyectoVinculacionEstudiante.objects.get(pk=id)
                    archivo = informe_estudiante.archivo.file
                    valido, msg, diccionario = verificarFirmasPDF(archivo)
                    return JsonResponse({'result': True,'context':diccionario})
                except Exception as ex:
                    return JsonResponse({'result': False, "mensaje": 'Error!: {}'.format(ex)}, safe=False)

            elif action == 'estudiantes_matriculados_vinculacion':
                try:
                    data['fecha'] = hoy = datetime.now()
                    data['nombre_reporte'] = nombre_reporte = 'ESTUDIANTES MATRICULADOS VINCULACIÓN - EN LÍNEA'
                    noti = Notificacion(cuerpo='Generación de reporte de excel en progreso',
                                        titulo=f'Reporte excel de {nombre_reporte}', destinatario=persona, url='', prioridad=1, app_label='SGA',
                                        fecha_hora_visible=datetime.now() + timedelta(days=1), tipo=2, en_proceso=True)
                    noti.save(request)
                    reporte_estudiantes_matriculados_vinculacion_background(request=request, data=data, notif=noti.pk, periodo=periodo).start()

                    return JsonResponse({"result": True,
                                         "mensaje": u"El reporte se está realizando. Verifique su apartado de notificaciones después de unos minutos.",
                                         "btn_notificaciones": traerNotificaciones(request, data, persona)})
                except Exception as ex:
                    return JsonResponse({'result': False, "mensaje": 'Error!: {}'.format(ex)})

            elif action == 'estudiantes_proceso_vinculacion':
                try:
                    data['fecha'] = hoy = datetime.now()
                    data['nombre_reporte'] = nombre_reporte = 'ESTUDIANTES PROCESO VINCULACIÓN'
                    noti = Notificacion(cuerpo='Generación de reporte de excel en progreso',
                                        titulo=f'Reporte excel de {nombre_reporte}', destinatario=persona, url='', prioridad=1, app_label='SGA',
                                        fecha_hora_visible=datetime.now() + timedelta(days=1), tipo=2, en_proceso=True)
                    noti.save(request)
                    reporte_estudiantes_proceso_vinculacion_background(request=request, data=data, notif=noti.pk, periodo=periodo).start()

                    return JsonResponse({"result": True,
                                         "mensaje": u"El reporte se está realizando. Verifique su apartado de notificaciones después de unos minutos.",
                                         "btn_notificaciones": traerNotificaciones(request, data, persona)})
                except Exception as ex:
                    return JsonResponse({'result': False, "mensaje": 'Error!: {}'.format(ex)})

            return HttpResponseRedirect(request.path)

        else:
            try:
                panel = None
                programasinvestigacion = ProgramasInvestigacion.objects.select_related().filter(status=True).order_by('-id')
                proyectosvinculacion = ProyectosInvestigacion.objects.select_related().filter(status=True, tipo=1).order_by('-aprobacion', '-programa__fecha_creacion', 'nombre')
                participantesmatrices = ParticipantesMatrices.objects.filter(status=True, matrizevidencia_id=2)
                data['experto_presu'] = experto_presu
                if 'panel' in request.GET:
                    panel = request.GET['panel']
                    if panel == '2':
                        filtros = Q(status=True, tipo=1)
                        data['estadoproyectos'] = ESTADOS_PROYECTO_VINCULACION_INVESTIGACION
                        search = None
                        ids = None
                        anio, cod, nombre, estado, docente, estudiante, url_vars = request.GET.get('anio', ''), request.GET.get('cod', ''), request.GET.get('nom', ''), request.GET.get('estado', ''), request.GET.get('doc', ''), request.GET.get('est', ''), ''
                        # if 'id' in request.GET:
                        #     ids = request.GET['id']
                        #     proyectos = proyectosvinculacion.filter(pk=ids, tipo=1).order_by('nombre')
                        # elif 'inscripcionid' in request.GET:
                        #     inscripcionid = request.GET['inscripcionid']
                        #     proyectos = Graduado.objects.filter(inscripcion__carrera__in=miscarreras).filter(inscripcion__id=inscripcionid)

                        if anio:
                            filtros = filtros & Q(fechainicio__year=anio)
                            data['anio'] = int(anio)
                            url_vars += "&anio={}".format(anio)
                        if cod:
                            filtros = filtros & Q(id=cod)
                            data['cod'] = int(cod)
                            url_vars += "&cod={}".format(cod)

                        if estudiante:
                            if ' ' in estudiante:
                                s = estudiante.split(" ")
                                proyectos_estudiantes = participantesmatrices.values_list('proyecto_id', flat=True).filter((
                                                                                         Q(inscripcion__persona__apellido1__contains=s[0])
                                                                                         & Q(inscripcion__persona__apellido2__contains=s[1])
                                                                                         ))
                            else:
                                proyectos_estudiantes = participantesmatrices.values_list('proyecto_id', flat=True).filter(
                                    Q(inscripcion__persona__nombres__contains=estudiante) |
                                    Q(inscripcion__persona__apellido1__contains=estudiante) |
                                    Q(inscripcion__persona__apellido2__contains=estudiante) |
                                    Q(inscripcion__persona__cedula__contains=estudiante)
                                )
                            filtros = filtros & Q(id__in=proyectos_estudiantes)
                            data['est'] = estudiante
                            url_vars += "&est={}".format(estudiante)

                        if docente:
                            if ' ' in docente:
                                s = docente.split(" ")
                                proyectos_docente = participantesmatrices.values_list('proyecto_id', flat=True).filter((
                                                                                         Q(profesor__persona__apellido1__contains=s[0])
                                                                                         & Q(profesor__persona__apellido2__contains=s[1])
                                                                                         ))
                            else:
                                proyectos_docente = participantesmatrices.values_list('proyecto_id', flat=True).filter(
                                    Q(profesor__persona__nombres__contains=docente) |
                                    Q(profesor__persona__apellido1__contains=docente) |
                                    Q(profesor__persona__apellido2__contains=docente) |
                                    Q(profesor__persona__cedula__contains=docente)
                                )
                            filtros = filtros & Q(id__in=proyectos_docente)
                            data['doc'] = docente
                            url_vars += "&doc={}".format(docente)

                        if estado:
                            if int(estado)>0:
                                data['estado'] = int(estado)
                                filtros = filtros & Q(aprobacion=estado)
                                url_vars += "&estado={}".format(estado)

                        if nombre:
                            filtros = filtros & Q(nombre__contains=nombre)
                            data['nom'] = nombre
                            url_vars += "&nom={}".format(nombre)

                        data["url_vars"] = url_vars
                        proyectos = proyectosvinculacion.filter(filtros).order_by('-aprobacion', '-programa__fecha_creacion', 'nombre')
                        paging = MiPaginador(proyectos, 25)
                        p = 1
                        try:
                            paginasesion = 1
                            if 'paginador' in request.session:
                                paginasesion = int(request.session['paginador'])
                            if 'page' in request.GET:
                                p = int(request.GET['page'])
                            else:
                                p = paginasesion
                            try:
                                page = paging.page(p)
                            except:
                                p = 1
                            page = paging.page(p)
                        except:
                            page = paging.page(p)
                        request.session['paginador'] = p
                        data['paging'] = paging
                        data['page'] = page
                        data['rangospaging'] = paging.rangos_paginado(p)
                        data['proyectos'] = page.object_list
                        data['search'] = search if search else ""
                        data['ids'] = ids if ids else ""

                        conv = FechaProyectos.objects.filter(status=True, fechainicio__lte=datetime.now(), fechafin__gte=datetime.now())
                        if conv.exists():
                            data['convocatoria'] = conv.last()
                            if not ProyectosInvestigacion.objects.filter(status=True, convocatoria=conv.last(), usuario_creacion=request.user).exists():
                                data['conv_activa'] = True
                    elif panel == '3':
                        data['evidencias'] = Evidencia.objects.filter(status=True, matrizevidencia_id=2)
                    elif panel == '4':
                        matriculas = Matricula.objects.filter(nivel__periodo_id=periodo.id,
                                                              status=True,

                                                              ).exclude(inscripcion__coordinacion__id__in=[9, 7, 10])

                        carreras = matriculas.values_list('inscripcion__carrera_id', flat=True)
                        data['totalMatricula'] = total = matriculas.count()
                        data['carreras'] = Carrera.objects.filter(pk__in = carreras).distinct()

                else:
                    search = None
                    ids = None
                    inscripcionid = None
                    if 'id' in request.GET:
                        ids = request.GET['id']
                        programas = programasinvestigacion.filter(inscripcion__carrera__in=miscarreras).filter(pk=ids).order_by('-id')
                    elif 's' in request.GET:
                        search = request.GET['s']
                        if search.isdigit():
                            programas = programasinvestigacion.filter(pk=search, status=True).order_by('-id')
                        else:
                            programas = programasinvestigacion.filter(nombre__icontains=search, status=True).order_by('-id')
                    else:
                        programas = programasinvestigacion.filter(status=True).order_by('-id')
                    paging = MiPaginador(programas, 25)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['page'] = page
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['programas'] = page.object_list
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['inscripcionid'] = inscripcionid if inscripcionid else ""
                    data['carreras'] = Carrera.objects.all().order_by('nombre')
                data['panel'] = panel
                data['cantidad_programas'] = programasinvestigacion.count()
                data['programas_activos'] = programasinvestigacion.filter(estado=True).count()
                data['programas_inactivos'] = programasinvestigacion.filter(estado=False).count()

                data['cantidad_proyectos'] = proyectosvinculacion.count()
                data['proyectos_aprobados'] = proyectosvinculacion.filter(aprobacion=1).count()

                return render(request, "inv_vinculacion/view.html", data)
            except Exception as ex:
                pass
