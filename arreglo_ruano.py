#!/usr/bin/env python
# -*- coding: utf-8 -*-
import os
import sys
import io

from django.core.exceptions import ObjectDoesNotExist
from django.http import HttpResponse
from openpyxl import load_workbook

SITE_ROOT = os.path.dirname(os.path.realpath(__file__))

# SITE_ROOT = os.path.dirname(os.path.realpath(__file__))
YOUR_PATH = os.path.dirname(os.path.realpath(__file__))
# print(f"YOUR_PATH: {YOUR_PATH}")
SITE_ROOT = os.path.join(SITE_ROOT, '')
# print(f"SITE_ROOT: {SITE_ROOT}")
your_djangoproject_home = os.path.split(SITE_ROOT)[0]
# print(f"your_djangoproject_home: {your_djangoproject_home}")
sys.path.append(your_djangoproject_home)
from django.core.wsgi import get_wsgi_application
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "settings")
application = get_wsgi_application()
from sga.models import Persona, HistoricoRecordAcademico, RecordAcademico, Materia
from idioma.models import GrupoInscripcion, Grupo, Periodo, PeriodoAsignatura, GrupoInscripcionAsignatura
from sagest.models import Pago
from django.contrib.auth.models import User
from sga.funciones import MiPaginador, puede_realizar_accion, log, convertir_lista
from inno.models import *
from sga.templatetags.sga_extras import encrypt
from urllib.request import urlopen, Request
from django.db import transaction, connection
import json
from xlwt import *
from xlwt import easyxf
import xlwt
import openpyxl
import pandas as pd

def asignacioncorreosnone():
    try:
        userids = Persona.objects.values_list('usuario_id', flat=True).filter(status=True, emailinst='none@unemi.edu.ec').exclude(usuario_id__isnull=True)
        for userid in userids:
            persona = Persona.objects.get(usuario_id=userid)
            usuario = User.objects.filter(id=userid).first()
            correoins = persona.emailinst
            nombre_usuario, dominio = correoins.split("@")
            nuevo_correo = f'{usuario.username}' + "@" + dominio
            print(nuevo_correo)
            persona.emailinst = nuevo_correo
            persona.save()

    except Exception as ex:
        print(ex)
        pass



def migrarmodulosinglesrecordacademico():
    try:
        grupoinscripcion = GrupoInscripcion.objects.filter(grupo_id__in=(1,2,3,4),observacion="",
                                                           status=True).order_by('inscripcion__persona_apellido1')
        contador=0
        for list in grupoinscripcion:
            idcursomoodle = list.grupo.idcursomoodle
            url = 'https://upei.buckcenter.edu.ec/usernamecoursetograde.php?username=%s&curso=%s' % (
                list.inscripcion.persona.identificacion(), idcursomoodle)
            print(url)
            req = Request(url)
            response = urlopen(req)
            result = json.loads(response.read().decode())
            idcurso = int(result['idcurso'])
            if result['nota'] != 'null':
                if idcurso == idcursomoodle:
                    if result['nota_up'] == "-":
                        nota = 0
                    else:
                        nota = result['nota_up']
                    list.nota = nota
                    observacion = result['estado']
                    list.observacion = observacion
                    list.save()
                list.obtener_creditos_horas_modulo()
                asignatura = None
                if list.nota == 70:
                    asignatura = PeriodoAsignatura.objects.filter(periodo_id=1,
                                                                  asignatura_id=783, status=True)
                elif list.nota > 70 and list.nota <= 80:
                    asignatura = PeriodoAsignatura.objects.filter(periodo_id=1,
                                                                  asignatura_id__in=(783, 784), status=True)
                elif list.nota > 80 and list.nota <= 90:
                    asignatura = PeriodoAsignatura.objects.filter(periodo_id=1,
                                                                  asignatura_id__in=(783, 784, 785), status=True)
                elif list.nota > 90 and list.nota <= 100:
                    asignatura = PeriodoAsignatura.objects.filter(periodo_id=1,
                                                                  asignatura_id__in=(783, 784, 785, 786),
                                                                  status=True)
                if asignatura:
                    for asig in asignatura:
                        list.estado = 1
                        list.save()
                        try:
                            eGrupoInscripcionAsignatura = GrupoInscripcionAsignatura.objects.get(
                                grupoinscripcion=list, asignatura=asig.asignatura)
                        except ObjectDoesNotExist:
                            eGrupoInscripcionAsignatura = GrupoInscripcionAsignatura(grupoinscripcion=list, asignatura=asig.asignatura)
                            eGrupoInscripcionAsignatura.save()
                        if not RecordAcademico.objects.filter(inscripcion=list.inscripcion, asignatura=asig.asignatura,
                                                              status=True):
                            eMateriaAsignadas = MateriaAsignada.objects.filter(materia__nivel__periodo_id=177,
                                                                               materia__asignatura=asig.asignatura,
                                                                               matricula__inscripcion=list.inscripcion,
                                                                               materia__nivel_id=1501,
                                                                               matricula__nivel__periodo_id=177)
                            if eMateriaAsignadas.values('id').exists():
                                erubros = eMateriaAsignadas.first().rubro.all()
                                if not Pago.objects.filter(rubro__in=erubros).exists():
                                    erubros.delete()
                                eMateriaAsignadas.delete()
                            if not RecordAcademico.objects.filter(inscripcion=list.inscripcion,
                                                     asignatura=asig.asignatura,
                                                     grupoinscripcion=list).exists():
                                record = RecordAcademico(inscripcion=list.inscripcion,
                                                     asignatura=asig.asignatura,
                                                     grupoinscripcion=list,
                                                     modulomalla=list.inscripcion.asignatura_en_modulomalla(
                                                         asig.asignatura),
                                                     nota=list.nota,
                                                     asistencia=100,
                                                     fecha=datetime.now().date(),
                                                     aprobada=True,
                                                     noaplica=False,
                                                     convalidacion=False,
                                                     pendiente=False,
                                                     creditos=list.obtener_creditos_horas_modulo()[0][1],
                                                     horas=list.obtener_creditos_horas_modulo()[0][0],
                                                     homologada=False,
                                                     valida=True,
                                                     validapromedio=False,
                                                     observaciones='%s'%list.grupo.periodo,
                                                     suficiencia=True)
                                record.save()
                            else:
                                record=RecordAcademico.objects.get(inscripcion=list.inscripcion,
                                                               asignatura=asig.asignatura,
                                                               grupoinscripcion=list)
                            if not record.historicorecordacademico_set.filter(status=True, fecha=record.fecha).exists():
                                nuevohistorico = HistoricoRecordAcademico(recordacademico=record,
                                                                          inscripcion=record.inscripcion,
                                                                          modulomalla=record.modulomalla,
                                                                          asignaturamalla=record.asignaturamalla,
                                                                          asignatura=record.asignatura if record.asignatura else record.asignaturamallahistorico.asignatura,
                                                                          grupoinscripcion=record.grupoinscripcion,
                                                                          nota=record.nota,
                                                                          asistencia=record.asistencia,
                                                                          sinasistencia=record.sinasistencia,
                                                                          fecha=record.fecha,
                                                                          noaplica=record.noaplica,
                                                                          aprobada=record.aprobada,
                                                                          convalidacion=record.convalidacion,
                                                                          homologada=record.homologada,
                                                                          pendiente=record.pendiente,
                                                                          creditos=record.creditos,
                                                                          horas=record.horas,
                                                                          valida=record.valida,
                                                                          validapromedio=record.validapromedio,
                                                                          materiaregular=record.materiaregular,
                                                                          materiacurso=record.materiacurso,
                                                                          observaciones=record.observaciones,
                                                                          completonota=record.completonota,
                                                                          completoasistencia=record.completoasistencia,
                                                                          suficiencia=record.suficiencia)
                                nuevohistorico.save()
                            seleccionada = \
                                record.historicorecordacademico_set.filter(status=True).order_by('-aprobada', '-fecha')[
                                    0]
                            record.asignaturamalla = seleccionada.asignaturamalla
                            record.asignatura = seleccionada.asignatura if record.asignatura else None
                            record.grupoinscripcion = seleccionada.grupoinscripcion
                            record.nota = seleccionada.nota
                            record.asistencia = seleccionada.asistencia
                            record.sinasistencia = seleccionada.sinasistencia
                            record.fecha = seleccionada.fecha
                            record.noaplica = seleccionada.noaplica
                            record.aprobada = seleccionada.aprobada
                            record.convalidacion = seleccionada.convalidacion
                            record.homologada = seleccionada.homologada
                            record.pendiente = seleccionada.pendiente
                            record.creditos = seleccionada.creditos
                            record.horas = seleccionada.horas
                            record.valida = seleccionada.valida
                            record.validapromedio = seleccionada.validapromedio
                            record.materiaregular = seleccionada.materiaregular
                            record.materiacurso = seleccionada.materiacurso
                            record.observaciones = seleccionada.observaciones
                            record.completonota = seleccionada.completonota
                            record.completoasistencia = seleccionada.completoasistencia
                            record.suficiencia = seleccionada.suficiencia
                            record.save()
                            list.inscripcion.actualizar_nivel()
                            list.inscripcion.actualiza_matriculas(
                                record.asignatura if record.asignatura else record.asignaturamallahistorico.asignatura)
                else:
                    list.estado = 2
                    list.save()
            contador =contador +1
            print("Registros exitosos: "+str(contador))
    except Exception as ex:
        print(ex)
        pass


def importarexportarexcel1():
    try:
        miarchivo = openpyxl.load_workbook("D:\Grupo_de_experimental_24Sept2022.xlsx")
        lista = miarchivo['Applab']
        totallista = lista.rows
        contador=0
        cursor = connection.cursor()
        for filas in totallista:
            contador += 1
            if contador >2:
                cedula = str(filas[27].value).strip()
                asignatura = str(filas[24].value).strip()
                sql = f"""SELECT 
                            carr.nombre,
                            nivmal.nombre,
                            EXTRACT(YEAR FROM NOW())-EXTRACT(YEAR FROM pers.nacimiento),
                            estcivil.nombre,
                            sex.nombre,
                            ins.colegio,
                            '' ,
                            '' ,
                            ins.puntajesenescyt,
                            ins.fechainicioprimernivel,
                            '' ,
                            (CASE 
                            WHEN perf.tienediscapacidad =FALSE THEN 'NO'
                            WHEN perf.tienediscapacidad =tRUE THEN 'SI'
                            END),
                            '',
                            sitlab.disponetrabajo,
                            sitlab.lugartrabajo,
                            (CASE 
                            WHEN ficha.escabezafamilia =FALSE THEN 'NO'
                            WHEN ficha.escabezafamilia =tRUE THEN 'SI'
                            END),
                            ocu.nombre,
                            (CASE WHEN ficha.estadogeneral=1 THEN 'Excelente'
                            WHEN ficha.estadogeneral=2 THEN 'Bueno'
                            WHEN ficha.estadogeneral=3 THEN 'Regular'
                             END),
                            '',
                            tipoviv.nombre,
                            tipovivpro.nombre,
                            ficha.val_usainternetseism,
                            '',
                            (CASE 
                            WHEN matr.becado =FALSE THEN 'NO'
                            WHEN matr.becado =tRUE THEN 'SI'
                            END),
                            pext.hijos,
                            ficha.enfermedadescomunes,
                            (CASE 
                            WHEN perdat.tienediscapacidad=FALSE THEN 'NO'
                            WHEN perdat.tienediscapacidad=TRUE THEN 'SI'
                            END)
                            
                            FROM sga_inscripcion ins 
                            inner JOIN sga_matricula matr ON matr.inscripcion_id=ins.id
                            INNER JOIN sga_materiaasignada matasig ON matasig.matricula_id=matr.id
                            INNER JOIN sga_tipoestado est ON est.id=matasig.estado_id
                            INNER JOIN sga_materia mate ON mate.id=matasig.materia_id
                            INNER JOIN sga_inscripcionnivel insniv ON insniv.inscripcion_id=ins.id
                            INNER JOIN sga_nivelmalla nivmal ON nivmal.id=insniv.nivel_id
                            INNER JOIN sga_asignatura asig ON mate.asignatura_id=asig.id
                            INNER JOIN sga_carrera carr ON carr.id=ins.carrera_id
                            INNER JOIN sga_coordinacion coor ON coor.id=ins.coordinacion_id
                            INNER JOIN sga_persona pers ON ins.persona_id=pers.id
                            left JOIN sga_personadatosfamiliares perdat ON perdat.persona_id=pers.id
                            --NUEVO
                            INNER JOIN socioecon_fichasocioeconomicainec ficha ON ficha.persona_id=pers.id
                            INNER JOIN socioecon_ocupacionjefehogar ocu ON ocu.id=ficha.ocupacionjefehogar_id
                            INNER JOIN socioecon_tipovivienda tipoviv ON tipoviv.id=ficha.tipovivienda_id
                            INNER JOIN socioecon_tipoviviendapro tipovivpro ON tipovivpro.id=ficha.tipoviviendapro_id
                            INNER JOIN sga_perfilinscripcion perf ON perf.persona_id=pers.id
                            LEFT JOIN sga_personasituacionlaboral sitlab ON sitlab.persona_id=pers.id
                            INNER JOIN sga_sexo sex ON sex.id=pers.sexo_id
                            INNER JOIN med_personaextension pext ON pers.id=pext.persona_id
                            INNER JOIN sga_personaestadocivil estcivil ON estcivil.id = pext.estadocivil_id
                            WHERE pers.cedula = '{cedula}' AND coor.id=4 AND asig.nombre ILIKE '%{asignatura}%'
                            """#%(cedula,asignatura)
                cursor.execute(sql)
                results = cursor.fetchall()
                for per in results:
                    filas[25].value = per[0]
                    filas[26].value = per[1]
                    filas[28].value = per[2]
                    filas[29].value = per[3]
                    filas[30].value = per[4]
                    filas[31].value = per[5]
                    filas[32].value = per[6]
                    filas[33].value = per[7]
                    filas[34].value = per[8]
                    filas[35].value = per[9]
                    filas[36].value = per[10]
                    filas[37].value = per[11]
                    filas[38].value = per[12]
                    filas[39].value = per[13]
                    filas[40].value = per[14]
                    filas[41].value = per[15]
                    filas[42].value = per[16]
                    filas[43].value = per[17]
                    filas[44].value = per[18]
                    filas[45].value = per[19]
                    filas[46].value = per[20]
                    filas[47].value = per[21]
                    filas[48].value = per[22]
                    filas[49].value = per[23]
                    filas[50].value = per[24]
                    filas[51].value = per[25]
                    filas[52].value = per[26]
        miarchivo.save("D:\Grupo_de_experimental_24Sept2022.xlsx")
        print(u"fin")

    except Exception as ex:
        print(sys.exc_info()[-1].tb_lineno)
        print(ex)
        pass


def importarexportarexcel2():
    try:
        miarchivo = openpyxl.load_workbook("D:\Grupo_de_control_24Sept2022.xlsx")
        lista = miarchivo['CALIFICACIONES LLEVADAS A 20']
        totallista = lista.rows
        contador = 0
        cursor = connection.cursor()
        for filas in totallista:
            contador += 1
            if contador > 2:
                cedula = str(filas[28].value).strip()
                asignatura = str(filas[25].value).strip()
                sql = f"""SELECT 
                            carr.nombre,
                            nivmal.nombre,
                            EXTRACT(YEAR FROM NOW())-EXTRACT(YEAR FROM pers.nacimiento),
                            estcivil.nombre,
                            sex.nombre,
                            ins.colegio,
                            '' ,
                            '' ,
                            ins.puntajesenescyt,
                            ins.fechainicioprimernivel,
                            '' ,
                            (CASE 
                            WHEN perf.tienediscapacidad =FALSE THEN 'NO'
                            WHEN perf.tienediscapacidad =tRUE THEN 'SI'
                            END),
                            '',
                            sitlab.disponetrabajo,
                            sitlab.lugartrabajo,
                            (CASE 
                            WHEN ficha.escabezafamilia =FALSE THEN 'NO'
                            WHEN ficha.escabezafamilia =tRUE THEN 'SI'
                            END),
                            ocu.nombre,
                            (CASE WHEN ficha.estadogeneral=1 THEN 'Excelente'
                            WHEN ficha.estadogeneral=2 THEN 'Bueno'
                            WHEN ficha.estadogeneral=3 THEN 'Regular'
                             END),
                            '',
                            tipoviv.nombre,
                            tipovivpro.nombre,
                            ficha.val_usainternetseism,
                            '',
                            (CASE 
                            WHEN matr.becado =FALSE THEN 'NO'
                            WHEN matr.becado =tRUE THEN 'SI'
                            END),
                            pext.hijos,
                            ficha.enfermedadescomunes,
                            (CASE 
                            WHEN perdat.tienediscapacidad=FALSE THEN 'NO'
                            WHEN perdat.tienediscapacidad=TRUE THEN 'SI'
                            END)

                            FROM sga_inscripcion ins 
                            inner JOIN sga_matricula matr ON matr.inscripcion_id=ins.id
                            INNER JOIN sga_materiaasignada matasig ON matasig.matricula_id=matr.id
                            INNER JOIN sga_tipoestado est ON est.id=matasig.estado_id
                            INNER JOIN sga_materia mate ON mate.id=matasig.materia_id
                            INNER JOIN sga_inscripcionnivel insniv ON insniv.inscripcion_id=ins.id
                            INNER JOIN sga_nivelmalla nivmal ON nivmal.id=insniv.nivel_id
                            INNER JOIN sga_asignatura asig ON mate.asignatura_id=asig.id
                            INNER JOIN sga_carrera carr ON carr.id=ins.carrera_id
                            INNER JOIN sga_coordinacion coor ON coor.id=ins.coordinacion_id
                            INNER JOIN sga_persona pers ON ins.persona_id=pers.id
                            left JOIN sga_personadatosfamiliares perdat ON perdat.persona_id=pers.id
                            --NUEVO
                            LEFT JOIN socioecon_fichasocioeconomicainec ficha ON ficha.persona_id=pers.id
                            LEFT JOIN socioecon_ocupacionjefehogar ocu ON ocu.id=ficha.ocupacionjefehogar_id
                            LEFT JOIN socioecon_tipovivienda tipoviv ON tipoviv.id=ficha.tipovivienda_id
                            LEFT JOIN socioecon_tipoviviendapro tipovivpro ON tipovivpro.id=ficha.tipoviviendapro_id
                            LEFT JOIN sga_perfilinscripcion perf ON perf.persona_id=pers.id
                            LEFT JOIN sga_personasituacionlaboral sitlab ON sitlab.persona_id=pers.id
                            INNER JOIN sga_sexo sex ON sex.id=pers.sexo_id
                            INNER JOIN med_personaextension pext ON pers.id=pext.persona_id
                            INNER JOIN sga_personaestadocivil estcivil ON estcivil.id = pext.estadocivil_id
                            WHERE pers.cedula = '{cedula}' AND coor.id=4 AND asig.nombre ILIKE '%{asignatura}%'
                            """  # %(cedula,asignatura)
                cursor.execute(sql)
                results = cursor.fetchall()
                for per in results:
                    filas[26].value = per[0] #CARRERA
                    filas[27].value = per[1] #NIVEL
                    filas[29].value = per[2] #EDAD
                    filas[30].value = per[3] #ESTADO CIVIL
                    filas[31].value = per[4] #SEXO
                    filas[32].value = per[5] #COLEGIO
                    filas[33].value = per[6] #TIPO COLEGIO
                    filas[34].value = per[7] #FECHA BACHILLER
                    filas[35].value = per[8] #ACCESOS PUNTAJE IES
                    filas[36].value = per[9] #FECHA INGRESO PRIMER NIVEL
                    filas[37].value = per[10] #PROBLEMA DE APRENDIZAJE
                    filas[38].value = per[11] #DISCAPACIDAD
                    filas[39].value = per[12] #TIPO DISCAPACIDAD
                    filas[40].value = per[13] #TIENE TRABAJO
                    filas[41].value = per[14] #LUGAR DE TRABAJO
                    filas[42].value = per[15] #CABEZA DE FAMILIA
                    filas[43].value = per[16] #OCUPACION CABEZA DE FAMILIA
                    filas[44].value = per[17] #ESTADO DE SALUD
                    filas[45].value = per[18] #ESTUDIA OTRA CARRERA
                    filas[46].value = per[19] #TIPO DE VIVIENDA
                    filas[47].value = per[20] #VIVIENDA
                    filas[48].value = per[21] #HORAS USO INTERNET
                    filas[49].value = per[22] #RECIBE AYUDA FAMILIAR
                    filas[50].value = per[23] #RECIBE BECA
                    filas[51].value = per[24] #NUMERO DE HIJOS
                    filas[52].value = per[25] #ENFERMEDAD FAMILIAR
                    filas[53].value = per[26] #DISCAPACIDAD FAMILIAR
        miarchivo.save("D:\Grupo_de_control_24Sept2022.xlsx")
        print(u"fin")

    except Exception as ex:
        print(sys.exc_info()[-1].tb_lineno)
        print(ex)
        pass


def actualizaridmoodle():
    with transaction.atomic():
        try:
            miarchivo = openpyxl.load_workbook("D:\matriculadosconcodmoodle.xlsx")
            lista = miarchivo['Hoja1']
            totallista = lista.rows
            contador = 0
            for filas in totallista:
                contador += 1
                if contador > 1:
                    idmateria = int(filas[1].value)
                    idcursomoodle = int(filas[2].value)
                    materias = Materia.objects.get(pk=idmateria, status=True)
                    materias.idcursomoodle = idcursomoodle
                    materias.save()
            print("Proceso completado")

        except Exception as ex:
            transaction.set_rollback(True)
            print(sys.exc_info()[-1].tb_lineno)
            print(ex)
            print(f"Error al actualizar las asistencias")
        pass

