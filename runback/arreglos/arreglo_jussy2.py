#!/usr/bin/env python
# -*- coding: utf-8 -*-
import os
import sys
import openpyxl



# SITE_ROOT = os.path.dirname(os.path.realpath(__file__))


YOUR_PATH = os.path.dirname(os.path.realpath(__file__))
# print(f"YOUR_PATH: {YOUR_PATH}")
SITE_ROOT = os.path.dirname(os.path.dirname(YOUR_PATH))
SITE_ROOT = os.path.join(SITE_ROOT, '')
# print(f"SITE_ROOT: {SITE_ROOT}")
your_djangoproject_home = os.path.split(SITE_ROOT)[0]
# print(f"your_djangoproject_home: {your_djangoproject_home}")
sys.path.append(your_djangoproject_home)

from django.core.wsgi import get_wsgi_application
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "settings")
application = get_wsgi_application()
from sga.models import *
from sagest.models import *
from inno.models import *
from matricula.models import *
from sga.commonviews import actualizar_nota_planificacion
# from balcon.models import *
from moodle import moodle
from urllib.request import urlopen, Request
import json
from sga.commonviews import ficha_socioeconomica
from Moodle_Funciones import crearhtmlphpmoodle, accesoQuizIndividual
from sga.funciones import convertir_hora,remover_caracteres_tildes_unicode,remover_caracteres_especiales_unicode
from certi.models import Carnet
from oma.models import Curso, AsignaturaInscripcionCurso, AuditoriaNotasOma, InscripcionCurso, AsignaturaCurso,\
    DetalleModeloEvaluativo as DetalleModeloEvaluativoOMA, EvaluacionGenerica as EvaluacionGenericaOMA
from oma.oma_curso import actualizar_nota_ofimatica, generarCertificadoOfimatica
from postulate.models import Programa as ProgramaPostulate, ArmonizacionNomenclaturaTitulo, \
    PartidaArmonizacionNomenclaturaTitulo, PartidaAsignaturas
from bd.models import LogQuery
from sga.funcionesxhtml2pdf import conviert_html_to_pdf, conviert_html_to_pdf_name, \
    conviert_html_to_pdfsaveqr_omacertificado
import warnings
warnings.filterwarnings('ignore', message='Unverified HTTPS request')
print(u"Inicio")
# cursos = Materia.objects.filter(nivel__periodo_id=136, asignaturamalla__nivelmalla_id=1,
#                                 asignaturamalla__malla__carrera_id=101,
#                                 status=True,
#                                 materiaasignada__status=True,
#                                 asignatura_id=4837).order_by('asignatura__nombre', 'inicio', 'identificacion', 'id').distinct()
# print(u"%s" % cursos)
# for curso in cursos:
#     print("entra")
#     curso.crear_actualizar_estudiantes_curso_admision(moodle, 2)
#     print(u"%s" % curso)




# try:
#     materias = Materia.objects.filter(status=True,nivel__periodo__id=119,asignaturamalla__malla__carrera__id__in=[139, 76])
#     persona=Persona.objects.get(id=22977)
#     fecha=convertir_fecha_hora("14-01-2022 00:00:00")
#     cont=0
#     for materia in materias:
#         cont+=1
#         for detalle in materia.detallemodeloevaluativo().filter(nombre="EX1"):
#             if ReactivoMateria.objects.filter(materia=materia, detallemodelo=detalle):
#                 reactivomateria=ReactivoMateria.objects.filter(materia=materia, detallemodelo=detalle)[0]
#                 reactivomateria.fecha=fecha
#                 print("%s Se modifica fecha %s "%(cont,reactivomateria))
#             else:
#                 reactivomateria = ReactivoMateria(materia=materia,
#                                                               fecha=fecha,
#                                                               persona=persona,
#                                                               detallemodelo=detalle)
#                 print("%s Se crea fecha %s" % (cont,reactivomateria))
#             reactivomateria.save()
# except Exception as ex:
#     print(ex)

# persona_responsable=Persona.objects.get(id=19)
# from Moodle_Funciones import CrearExamenMoodle
# fecha = int(time.mktime(datetime.now().timetuple()))
# # for test in TestSilaboSemanal.objects.filter(status=True, id=21592 ).distinct():
# for test in TestSilaboSemanal.objects.filter(status=True, estado_id=4, silabosemanal__examen=True,
#                                              silabosemanal__silabo__materia__nivel__periodo__id=119 ).distinct():
#     try:
#         conexion = connections['moodle_db']
#         cursor = conexion.cursor()
#         cursoid = test.silabosemanal.silabo.materia.idcursomoodle
#         sql = """select id from mooc_course_sections WHERE course=%s AND SECTION='%s' """ % (cursoid, 12)
#         cursor.execute(sql)
#         buscar = cursor.fetchall()
#         section = buscar[0][0]
#
#         instanceid=test.idtestmoodle
#         sql = """update mooc_course_modules set section='%s' where course=%s and module=17 and instance=%s""" % (
#             section, cursoid, instanceid)
#         cursor.execute(sql)
#
#         # sql = """select array_to_string(array_agg(id),',') from mooc_course_modules where course=%s and module=17 and instance in (select id from mooc_quiz where course=%s)""" % (cursoid, cursoid)
#         sql = """select array_to_string(array_agg(id),',') from mooc_course_modules where deletioninprogress=0 and course=%s and section=%s""" % (
#         cursoid, section)
#         cursor.execute(sql)
#         buscar = cursor.fetchall()
#         course_modules = buscar[0][0]
#
#         sql = """UPDATE mooc_course_sections SET sequence = '%s' WHERE course = %s and section=12""" % (
#         course_modules, cursoid)
#         cursor.execute(sql)
#
#         sql = """update mooc_course_modules set section='%s' where course=%s and module=17 and instance=%s""" % (
#         section, cursoid, instanceid)
#         cursor.execute(sql)
#
#         sql = u"UPDATE mooc_course SET cacherev = %s WHERE id = %s" % (fecha, cursoid)
#         cursor.execute(sql)
#
#         print(u"%s"%test)
#     except Exception as ex:
#         print('error: %s' % (ex))
#         pass
# print(u"Fin")


# for matricula in Matricula.objects.filter(nivel__periodo_id=119, status=True).exclude(estado_matricula=2):
    # for materiaasignada in matricula.mis_materias_sin_retiro():
    #     materiaasignada.materia.crear_actualizar_un_estudiante_curso(moodle, 1, matricula)
    #     print(u"%s" % materiaasignada)
    # matricula.estado_matricula=2
    # matricula.save()








# cursor = connections['moodle_db'].cursor()
# periodo = Periodo.objects.get(pk=119)
# LogDeberes.objects.filter(periodo=periodo).delete()
# for materia in Materia.objects.filter(nivel__periodo=periodo, status=True, idcursomoodle__gt=0,
#                                       asignaturamalla__malla__carrera__coordinacion__id__in=[1,2,3,4,5]).distinct():
#     cursomoodle = materia.idcursomoodle
#     profesor = materia.profesor_principal()
#     revisor=None
#     if materia.profesormateria_set.values('id').filter(status=True, activo=True, tipoprofesor__id=8).exists():
#         revisor = materia.profesormateria_set.filter(status=True, activo=True, tipoprofesor__id=8)[0].profesor.persona
#     estudiantes = materia.cantidad_matriculas_materia()
#     sql = "SELECT b.name, u.idnumber FROM mooc_assign AS b INNER JOIN mooc_assign_grades AS c ON c.ASSIGNMENT=b.id " \
#           " INNER JOIN mooc_user as u ON c.grader=u.id WHERE b.course=" + str(
#         cursomoodle) + " GROUP by b.name, u.idnumber"
#     cursor.execute(sql)
#     results = cursor.fetchall()
#     cedula = ''
#     for r in results:
#         # if cedula != str(r[1]):
#         #     cedula = str(r[1])
#         #     revisoraux = Persona.objects.filter(Q(cedula=str(r[1])) or Q(pasaporte=str(r[1])))
#         #     revisor = None
#         #     if revisoraux:
#         #         revisor = revisoraux[0]
#         sql1 = "SELECT TO_TIMESTAMP(c.timemodified) FROM mooc_assign AS b " \
#                " INNER JOIN mooc_assign_grades AS c ON c.ASSIGNMENT=b.id WHERE b.course=" + str(
#             cursomoodle) + " and b.name='" + str(r[0]) + "' order by b.name, c.timemodified"
#         cursor.execute(sql1)
#         resultaux = cursor.fetchall()
#         # esto me ayudara a sacar la desviacion estandar
#
#         cantidad = resultaux.__len__()
#         suma = 0
#         i = 0
#         acumulador = 0
#         minimo = 500
#         maximo = 0
#         arreglo = []
#         while i < cantidad - 1:
#             fecha = (resultaux[i][0]).date()
#             tiempo = resultaux[i][0]
#             tiemposiguiente = resultaux[i + 1][0]
#             i += 1
#             restar = (tiemposiguiente - tiempo).seconds
#             if restar <= 300:
#                 acumulador = acumulador + restar
#             else:
#                 restar = 300
#                 acumulador = acumulador + restar
#             if minimo > restar:
#                 minimo = restar
#             if maximo < restar:
#                 maximo = restar
#             arreglo.append(restar)
#
#         l = LogDeberes(periodo=periodo,
#                        profesor=profesor,
#                        revisor=revisor,
#                        materia=materia,
#                        deber=str(r[0]),
#                        estudiantes=estudiantes,
#                        tiempo=acumulador,
#                        tiempominimo=minimo,
#                        tiempomaximo=maximo,
#                        fecha=fecha)
#         l.save()
#         print(u"%s"%lw



# import xlwt
# from xlwt import *
# from django.http import HttpResponse
#
# response = HttpResponse(content_type="application/ms-excel")
# response['Content-Disposition'] = 'attachment; filename=tiempo_calificacion.xls'
# style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
# style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
# style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
# title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
# style1 = easyxf(num_format_str='D-MMM-YY')
# font_style = XFStyle()
# font_style.font.bold = True
# font_style2 = XFStyle()
# font_style2.font.bold = False
# wb = xlwt.Workbook()
# ws = wb.add_sheet('Sheetname')
# estilo = xlwt.easyxf('font: height 350, name Arial, colour_index black, bold on, italic on; align: wrap on, vert centre, horiz center;')
# ws.write_merge(0, 0, 0, 9, 'UNIVERSIDAD ESTATAL DE MILAGRO', estilo)
# output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
# nombre = "Lista" + datetime.now().strftime('%Y%m%d_%H%M%S') + ".xls"
# filename = os.path.join(output_folder, nombre)
# columns = [(u"Docente", 6000),
#            (u"Revisor", 6000),
#            (u"Materia", 6000),
#            (u"Deber", 6000),
#            (u"Matriculados", 6000),
#            (u"Tiempo", 6000),
#            (u"Tiempo mínimo", 6000),
#            (u"Tiempo máximo", 6000),
#            ]
# row_num = 3
# for col_num in range(len(columns)):
#     ws.write(row_num, col_num, columns[col_num][0], font_style)
#     ws.col(col_num).width = columns[col_num][1]
# row_num = 4
#
# for l in LogDeberes.objects.filter(periodo_id=119):
#     try:
#         ws.write(row_num, 0, u'%s' % l.profesor, font_style2)
#         ws.write(row_num, 1, u'%s' % l.revisor, font_style2)
#         ws.write(row_num, 2, u'%s' % l.materia, font_style2)
#         ws.write(row_num, 3, u'%s' % l.deber, font_style2)
#         ws.write(row_num, 4, u'%s' % l.estudiantes, font_style2)
#         ws.write(row_num, 5, u'%s' % l.hora(), font_style2)
#         ws.write(row_num, 6, u'%s' % l.horaminima(), font_style2)
#         ws.write(row_num, 7, u'%s' % l.horamaxima(), font_style2)
#         row_num+=1
#         print('%s' % (l))
#     except Exception as ex:
#         print('error: %s' % (ex))
#         pass
# wb.save(filename)
# print("FIN: ", filename)

# ma=MateriaAsignada.objects.get(id=1686924)
# def calculo_modelo_evaluativo(ma):
# 	P1 = ma.campo('P1')
# 	N1 = ma.campo('N1')
# 	N2 = ma.campo('N2')
# 	EX1 = ma.campo('EX1')
# 	P2 = ma.campo('P2')
# 	N3 = ma.campo('N3')
# 	N4 = ma.campo('N4')
# 	EX2 = ma.campo('EX2')
# 	RE = ma.campo('RE')
# 	P1.valor=N1.valor + N2.valor + EX1.valor
# 	P1.save()
# 	P2.valor=N3.valor + N4.valor + EX2.valor
# 	P2.save()
# 	promedio = P1.valor + P2.valor
# 	ma.notafinal = null_to_decimal(promedio, 0)
# 	if ma.notafinal< 40:
# 		RE.valor = 0
# 		RE.save()
# 	elif ma.notafinal< 70:
# 		if RE.valor > 0:
# 			ma.notafinal = null_to_decimal((RE.valor + float(ma.notafinal)) / 2, 0)
# 	else:
# 		RE.valor = 0
# 		RE.save()
# 	if EX2.valor > 0 or RE.valor > 0:
# 		if not ma.sinasistencia:
# 			if ma.asistenciafinal < 70:
# 				EX2.valor = 0
# 				EX2.save()
# 				RE.valor = 0
# 				RE.save()
# 				P2.valor=N3.valor + N4.valor + EX2.valor
# 				P2.save()
# 				promedio = P1.valor + P2.valor
# 				ma.notafinal = null_to_decimal(promedio, 0)
# 	ma.save()

#
# instructor = CapInstructorIpec.objects.get(status=True, pk=272)
# grupo = instructor.capeventoperiodo
# codigointegrante = grupo.capinscritoipec_set.get(pk=14066, status=True)
# codigointegrante.encursomoodle = True
# codigointegrante.save()
# instructor.crear_curso_moodle(14066, 1)
# print("FIN")


# import xlwt
# from xlwt import *
# from django.http import HttpResponse
#
# response = HttpResponse(content_type="application/ms-excel")
# response['Content-Disposition'] = 'attachment; filename=tiempo_calificacion.xls'
# style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
# style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
# style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
# title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
# style1 = easyxf(num_format_str='D-MMM-YY')
# font_style = XFStyle()
# font_style.font.bold = True
# font_style2 = XFStyle()
# font_style2.font.bold = False
# wb = xlwt.Workbook()
# ws = wb.add_sheet('Sheetname')
# estilo = xlwt.easyxf('font: height 350, name Arial, colour_index black, bold on, italic on; align: wrap on, vert centre, horiz center;')
# ws.write_merge(0, 0, 0, 9, 'UNIVERSIDAD ESTATAL DE MILAGRO', estilo)
# output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
# nombre = "Lista" + datetime.now().strftime('%Y%m%d_%H%M%S') + ".xls"
# filename = os.path.join(output_folder, nombre)
# periodo=Periodo.objects.get(id=126)
# columns = [
#            (u"id malla", 9000),
#            (u"CARRERA", 9000),
#            (u"id carrera", 9000),
#            (u"FECHA INICIO", 9000),
#            (u"HORAS TOTALES", 6000),
#            (u"CREDITO TOTALES", 6000),
#            (u"NIVELES", 6000),
#            (u"NIVEL", 6000),
#            (u"CREDITO POR NIVEL", 6000),
#            (u"HORAS POR NIVEL", 6000),
#            ]
# row_num = 3
# for col_num in range(len(columns)):
#     ws.write(row_num, col_num, columns[col_num][0], font_style)
#     ws.col(col_num).width = columns[col_num][1]
# row_num = 4
#
# for malla in Malla.objects.filter(carrera__coordinacion__id__in=[1,2,3,4,5]).distinct():
#     try:
#         if malla.uso_en_periodo(periodo):
#             for nivel in malla.niveles_malla():
#                 ws.write(row_num, 0, u'%s' % malla.id, font_style2)
#                 ws.write(row_num, 1, u'%s' % malla.carrera.nombre_completo(), font_style2)
#                 ws.write(row_num, 2, u'%s' % malla.carrera.id, font_style2)
#                 ws.write(row_num, 3, u'%s' % malla.inicio, font_style2)
#                 ws.write(row_num, 4, u'%s' % malla.suma_horas_validacion_itinerario(), font_style2)
#                 ws.write(row_num, 5, u'%s' % malla.suma_creditos_validacion_itinerario(), font_style2)
#                 ws.write(row_num, 6, u'%s' % malla.cantidad_niveles(), font_style2)
#                 ws.write(row_num, 7, u'%s' % nivel.nombre, font_style2)
#                 ws.write(row_num, 8, u'%s' % nivel.suma_creditos_validacion_itinerario(malla), font_style2)
#                 ws.write(row_num, 9, u'%s' % nivel.suma_horas_validacion_itinerario(malla), font_style2)
#                 row_num+=1
#                 print('%s' % (row_num))
#     except Exception as ex:
#         print('error: %s' % (ex))
#         pass
# wb.save(filename)
# print("FIN: ", filename)

# from inno.models import ConfiguracionCalculoMatriculaNivel,DetalleConfiguracionCalculoMatriculaNivel
#
# try:
#     miarchivo = openpyxl.load_workbook("NUEVO CALCULO DE VALORES POR CARRERA.xlsx")
#     lista = miarchivo.get_sheet_by_name('calculo final x niveles')
#     totallista = lista.rows
#     a=0
#     periodo=Periodo.objects.get(id=126)
#     for filas in totallista:
#         a += 1
#         if a > 2:
#             id_malla = filas[0].value
#             costooptimo = filas[2].value
#             id_nivel = filas[8].value
#             vct = filas[12].value
#             valormatricula = filas[13].value
#             vctalto = filas[14].value
#             vctmedio = filas[15].value
#             vctmediotipico = filas[16].value
#             vctmediobajo = filas[17].value
#             vctbajo = filas[18].value
#             configuracion=None
#             if not ConfiguracionCalculoMatriculaNivel.objects.filter(status=True, malla_id=id_malla,periodo=periodo).exists():
#                 configuracion=ConfiguracionCalculoMatriculaNivel(status=True, malla_id=id_malla,periodo=periodo)
#                 configuracion.save()
#             else:
#                 configuracion=ConfiguracionCalculoMatriculaNivel.objects.filter(status=True, malla_id=id_malla, periodo=periodo)[0]
#             print('Registro ingresado cabecera: %s' % configuracion)
#             if not DetalleConfiguracionCalculoMatriculaNivel.objects.filter(status=True,configuracion=configuracion,gruposocioeconomico_id=1,nivel_id=id_nivel ).exists():
#                 detalle1=DetalleConfiguracionCalculoMatriculaNivel(gruposocioeconomico_id=1, configuracion=configuracion,
#                                                                    vct=vct,nivel_id=id_nivel,
#                                                                    valormatricula=valormatricula,
#                                                                    valornivelsocioeconomico=vctalto)
#                 detalle1.save()
#             else:
#                 detalle1=DetalleConfiguracionCalculoMatriculaNivel.objects.filter(status=True, configuracion=configuracion, gruposocioeconomico_id=1)[0]
#                 detalle1.vct = vct
#                 detalle1.nivel_id = id_nivel
#                 detalle1.valormatricula = valormatricula
#                 detalle1.valornivelsocioeconomico = vctalto
#                 detalle1.save()
#             print('Registro ingresado detalle alto: %s' % detalle1)
#
#             if not DetalleConfiguracionCalculoMatriculaNivel.objects.filter(status=True, configuracion=configuracion,
#                                                                             gruposocioeconomico_id=2, nivel_id=id_nivel).exists():
#                 detalle2 = DetalleConfiguracionCalculoMatriculaNivel(gruposocioeconomico_id=2, configuracion=configuracion,
#                                                                      vct=vct, nivel_id=id_nivel,
#                                                                      valormatricula=valormatricula,
#                                                                      valornivelsocioeconomico=vctmedio)
#                 detalle2.save()
#             else:
#                 detalle2 = DetalleConfiguracionCalculoMatriculaNivel.objects.filter(status=True, configuracion=configuracion,
#                                                                                     gruposocioeconomico_id=2)[0]
#                 detalle2.vct = vct
#                 detalle2.nivel_id = id_nivel
#                 detalle2.valormatricula = valormatricula
#                 detalle2.valornivelsocioeconomico = vctmedio
#                 detalle2.save()
#             print('Registro ingresado detalle medio alto: %s' % detalle2)
#
#
#             if not DetalleConfiguracionCalculoMatriculaNivel.objects.filter(status=True, configuracion=configuracion,
#                                                                             gruposocioeconomico_id=3, nivel_id=id_nivel).exists():
#                 detalle3 = DetalleConfiguracionCalculoMatriculaNivel(gruposocioeconomico_id=3, configuracion=configuracion,
#                                                                      vct=vct, nivel_id=id_nivel,
#                                                                      valormatricula=valormatricula,
#                                                                      valornivelsocioeconomico=vctmediotipico)
#                 detalle3.save()
#             else:
#                 detalle3 = DetalleConfiguracionCalculoMatriculaNivel.objects.filter(status=True, configuracion=configuracion,
#                                                                                     gruposocioeconomico_id=3)[0]
#                 detalle3.vct = vct
#                 detalle3.nivel_id = id_nivel
#                 detalle3.valormatricula = valormatricula
#                 detalle3.valornivelsocioeconomico = vctmediotipico
#                 detalle3.save()
#             print('Registro ingresado detalle medio tipico: %s' % detalle3)
#
#
#             if not DetalleConfiguracionCalculoMatriculaNivel.objects.filter(status=True, configuracion=configuracion,
#                                                                             gruposocioeconomico_id=4, nivel_id=id_nivel).exists():
#                 detalle4 = DetalleConfiguracionCalculoMatriculaNivel(gruposocioeconomico_id=4, configuracion=configuracion,
#                                                                      vct=vct, nivel_id=id_nivel,
#                                                                      valormatricula=valormatricula,
#                                                                      valornivelsocioeconomico=vctmediobajo)
#                 detalle4.save()
#             else:
#                 detalle4 = DetalleConfiguracionCalculoMatriculaNivel.objects.filter(status=True, configuracion=configuracion,
#                                                                                     gruposocioeconomico_id=4)[0]
#                 detalle4.vct = vct
#                 detalle4.nivel_id = id_nivel
#                 detalle4.valormatricula = valormatricula
#                 detalle4.valornivelsocioeconomico = vctmediobajo
#                 detalle4.save()
#             print('Registro ingresado detalle medio bajo: %s' % detalle4)
#
#             if not DetalleConfiguracionCalculoMatriculaNivel.objects.filter(status=True, configuracion=configuracion,
#                                                                             gruposocioeconomico_id=5, nivel_id=id_nivel).exists():
#                 detalle5 = DetalleConfiguracionCalculoMatriculaNivel(gruposocioeconomico_id=5, configuracion=configuracion,
#                                                                      vct=vct, nivel_id=id_nivel,
#                                                                      valormatricula=valormatricula,
#                                                                      valornivelsocioeconomico=vctbajo)
#                 detalle5.save()
#             else:
#                 detalle5 = DetalleConfiguracionCalculoMatriculaNivel.objects.filter(status=True, configuracion=configuracion,
#                                                                                     gruposocioeconomico_id=5)[0]
#                 detalle5.vct = vct
#                 detalle5.nivel_id = id_nivel
#                 detalle5.valormatricula = valormatricula
#                 detalle5.valornivelsocioeconomico = vctbajo
#                 detalle5.save()
#             print('Registro ingresado detalle bajo: %s' % detalle5)
# except Exception as ex:
#         print('error: %s' % ex)

# inscripcion=Inscripcion.objects.get(id=80707)
# num_matriculas = 3
# malla = inscripcion.malla_inscripcion().malla
# for record in inscripcion.recordacademico_set.filter(status=True, aprobada=False, asignatura_id__in=[x.asignatura_id for x in
#                                                                                               malla.asignaturamalla_set.all()]):
#     cant=len(inscripcion.historicorecordacademico_set.values('id').filter(asignatura_id=record.asignatura.id).exclude(
#         materiaregular__nivel__periodo__cuentavecesmatricula=False))
#     print("%s"%cant)
#     if record.inscripcion.total_record_asignatura(record.asignatura) >= num_matriculas:
#         print("tiene mas de 3")
# print("NO tiene mas de 3")
#


# for instructor in CapInstructorIpec.objects.filter(id__in=[275]):
#     curso = instructor.capeventoperiodo
#     if curso.costo:
#         listadocodigo = curso.list_inscritos_costo()
#     else:
#         listadocodigo = curso.list_inscritos_sin_costo()
#     contador=0
#     for inscrito in listadocodigo.filter(id__in=[14509] ):
#         instructor.crear_actualizar_estudiantes_curso(moodle, 1, inscrito.id)
#         contador+=1



# import xlwt
# from xlwt import *
# from django.http import HttpResponse
#
# response = HttpResponse(content_type="application/ms-excel")
# response['Content-Disposition'] = 'attachment; filename=deudores_todo_matricula.xls'
# style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
# style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
# style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
# title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
# style1 = easyxf(num_format_str='D-MMM-YY')
# font_style = XFStyle()
# font_style.font.bold = True
# font_style2 = XFStyle()
# font_style2.font.bold = False
# wb = xlwt.Workbook()
# ws = wb.add_sheet('Sheetname')
# estilo = xlwt.easyxf('font: height 350, name Arial, colour_index black, bold on, italic on; align: wrap on, vert centre, horiz center;')
# ws.write_merge(0, 0, 0, 9, 'UNIVERSIDAD ESTATAL DE MILAGRO', estilo)
# output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
# nombre = "Lista" + datetime.now().strftime('%Y%m%d_%H%M%S') + ".xls"
# filename = os.path.join(output_folder, nombre)
# columns = [(u"N.", 6000),
#            (u"PERIODO", 6000),
#            (u"CEDULA", 6000),
#            (u"APELLIDOS Y NOMBRES", 6000),
#            (u"SEXO", 6000),
#            (u"EMAIL", 6000),
#            (u"EMAILINST", 6000),
#            (u"FACULTAD", 6000),
#            (u"CARRERA", 6000),
#            (u"TELEFONO", 6000),
#            (u"PAIS", 6000),
#            (u"PROVINCIA", 6000),
#            (u"CANTON", 6000),
#            (u"DIRECCION", 6000),
#            (u"ESTADO SOCIO ECONOMICO", 6000),
#            (u"TOTAL", 6000)
#            ]
# row_num = 3
# for col_num in range(len(columns)):
#     ws.write(row_num, col_num, columns[col_num][0], font_style)
#     ws.col(col_num).width = columns[col_num][1]
# row_num = 4
# periodo=Periodo.objects.get(id=126)
# matriculados = periodo.lista_total_matriculados_con_deuda()
# for matri in matriculados:
#     try:
#         auxiliarenrolar=0
#         if matri.estado_matricula == 3:
#             rubros_cancelados_matricula = Rubro.objects.values_list('id', flat=True).filter(status=True,
#                                                                                             matricula=matri,
#                                                                                             cancelado=True,
#                                                                                             tipo_id=2924)
#             rubros_cancelados_arancel = Rubro.objects.values_list('id', flat=True).filter(status=True,
#                                                                                           matricula=matri,
#                                                                                           cancelado=True,
#                                                                                           tipo_id=2923)
#             if len(rubros_cancelados_matricula) == 0 or len(rubros_cancelados_arancel) == 0:
#                 auxiliarenrolar = 1
#         else:
#             auxiliarenrolar = 1
#         if auxiliarenrolar == 1:
#             nombreperiodo = matri.nivel.periodo.nombre
#             fichasoc = matri.inscripcion.persona.fichasocioeconomicainec_set.filter(status=True)[0]
#             facultad = ''
#             if matri.inscripcion.coordinacion.nombre:
#                 facultad = matri.inscripcion.coordinacion.nombre
#             totalpagar = 0.0
#             if Rubro.objects.filter(status=True, persona=matri.inscripcion.persona, cancelado=False,
#                                     matricula__nivel__periodo=periodo).exists():
#                 totalpagar = null_to_decimal(
#                     Rubro.objects.filter(status=True, persona=matri.inscripcion.persona, cancelado=False,
#                                          matricula__nivel__periodo=periodo).aggregate(
#                         totalpagar=Sum('saldo'))['totalpagar'], 2)
#             ws.write(row_num, 0, row_num - 4)
#             ws.write(row_num, 1, nombreperiodo)
#             ws.write(row_num, 2, matri.inscripcion.persona.cedula)
#             ws.write(row_num, 3, matri.inscripcion.persona.apellido1 + ' ' + matri.inscripcion.persona.apellido2 + ' ' + matri.inscripcion.persona.nombres)
#             alu_sexo = matri.inscripcion.persona.sexo.nombre if matri.inscripcion.persona.sexo else 'SIN DEFINIR'
#             ws.write(row_num, 4, alu_sexo)
#             ws.write(row_num, 5, matri.inscripcion.persona.email)
#             ws.write(row_num, 6, matri.inscripcion.persona.emailinst)
#             ws.write(row_num, 7, facultad)
#             ws.write(row_num, 8, str(matri.inscripcion.carrera))
#             ws.write(row_num, 9, matri.inscripcion.persona.telefono)
#             ws.write(row_num, 10, matri.inscripcion.persona.pais.nombre if matri.inscripcion.persona.pais_id else "")
#             ws.write(row_num, 11,  matri.inscripcion.persona.provincia.nombre if matri.inscripcion.persona.provincia_id else "")
#             ws.write(row_num, 12, matri.inscripcion.persona.canton.nombre if matri.inscripcion.persona.canton_id else "")
#             ws.write(row_num, 13, matri.inscripcion.persona.direccion + ' ' + matri.inscripcion.persona.direccion2)
#             alu_grupoeco = fichasoc.grupoeconomico.nombre if fichasoc.grupoeconomico else 'NINGUNO'
#             alu_grupoecocod = fichasoc.grupoeconomico.codigo if fichasoc.grupoeconomico else ''
#             ws.write(row_num, 14, "{} {}".format(alu_grupoecocod, alu_grupoeco))
#             ws.write(row_num, 15, totalpagar)
#             row_num+=1
#             print("%s: "%row_num)
#     except Exception as ex:
#         print('error: %s' % (ex))
#         pass
# wb.save(filename)
# print("FIN: ", filename)

# materia=Materia.objects.get(id=47309  )
# matricula=Matricula.objects.get(pk=464052)
# materia.crear_actualizar_un_estudiante_curso(moodle,1, matricula)
# conexion = connections['moodle_db']
# cursor = conexion.cursor()
# periodo=Periodo.objects.get(id=126)
# try:
#     miarchivo = openpyxl.load_workbook("estudiantes eliminados moodle.xlsx")
#     lista = miarchivo.get_sheet_by_name('resultados')
#     totallista = lista.rows
#     a=0
#     for filas in totallista:
#         a += 1
#         if a > 1:
#             firstname = str(filas[0].value)
#             lasttname = str(filas[1].value)
#             sql = f" SELECT id FROM mooc_user WHERE lastname LIKE '%{lasttname}%' AND firstname LIKE '%{firstname}%' ORDER BY id asc"
#             cursor.execute(sql)
#             registros = cursor.fetchall()
#             print(u"%s" % len(registros))
#             ultimo=None
#             for x in registros:
#                 ultimo=x[0]
#             if ultimo:
#                 sqlconsultausuario = f" SELECT username FROM mooc_user WHERE id=%s"%ultimo
#                 cursor.execute(sqlconsultausuario)
#                 registrosuser = cursor.fetchall()
#                 if registrosuser:
#                     username = registrosuser[0][0]
#                     if Persona.objects.filter(usuario__username=username).exists():
#                         persona = Persona.objects.get(usuario__username=username)
#                         if Matricula.objects.filter(status=True, nivel__periodo=periodo, inscripcion__persona=persona,retiradomatricula=False).exists():
#                             matricula=Matricula.objects.filter(status=True, nivel__periodo=periodo, inscripcion__persona=persona)[0]
#                             for materiaasignada in MateriaAsignada.objects.filter(status=True, matricula=matricula,retiramateria=False):
#                                 materiaasignada.materia.crear_actualizar_un_estudiante_curso(moodle,1,matricula)
#                 print(u"ejecuto")
#
# except Exception as ex:
#         print('error: %s' % ex)




# try:
#     miarchivo = openpyxl.load_workbook("id zoom.xlsx")
#     lista = miarchivo.get_sheet_by_name('Hoja1')
#     totallista = lista.rows
#     a=0
#     for filas in totallista:
#         a += 1
#         if a > 1:
#             correo = str(filas[0].value)
#             idzoom = str(filas[1].value)
#             urlzoom=u"https://unemi-edu-ec.zoom.us/j/%s"%idzoom
#             print(u"%s"%correo)
#             if Persona.objects.filter(status=True,emailinst=correo.replace(" ","")).exists():
#                 persona = Persona.objects.get(status=True,emailinst=correo)
#                 if Profesor.objects.filter(status=True,persona=persona).exists():
#                     profesor=Profesor.objects.get(status=True,persona=persona)
#                     if profesor.urlzoom!=urlzoom:
#                         profesor.urlzoom=urlzoom
#                         profesor.save()
#                         print(u"ejecuto %s"%idzoom)
#                     else:
#                         print(u"NO ES DIFERENTE")
#                 else:
#                     print(u" NO HAY PROFESOR")
#             else:
#                 print(u"-NO HAY PERSONA")
# except Exception as ex:
#         print('error: %s' % ex)



# import xlwt
# from xlwt import *
# from django.http import HttpResponse
#
# response = HttpResponse(content_type="application/ms-excel")
# response['Content-Disposition'] = 'attachment; filename=encuesta_sedeexamen.xls'
# style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
# style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
# style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
# title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
# style1 = easyxf(num_format_str='D-MMM-YY')
# font_style = XFStyle()
# font_style.font.bold = True
# font_style2 = XFStyle()
# font_style2.font.bold = False
# wb = xlwt.Workbook()
# ws = wb.add_sheet('Sheetname')
# estilo = xlwt.easyxf('font: height 350, name Arial, colour_index black, bold on, italic on; align: wrap on, vert centre, horiz center;')
# ws.write_merge(0, 0, 0, 9, 'UNIVERSIDAD ESTATAL DE MILAGRO', estilo)
# output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
# nombre = "Lista" + datetime.now().strftime('%Y%m%d_%H%M%S') + ".xls"
# filename = os.path.join(output_folder, nombre)
# encuesta=EncuestaGrupoEstudiantes.objects.get(id=24)
# preguntas = encuesta.preguntaencuestagrupoestudiantes_set.filter(status=True).order_by('orden')
# columns = [ (u"Nº.", 2000),
#                 (u"CÉDULA", 3000),
#                 (u"ENCUESTADO", 9000),
#             ]
#
# if encuesta.tipoperfil == 1:  # ALUMNO
#     columns.append((u'CARRERA', 9000), )
#     columns.append((u'NUM MATERIAS', 9000), )
#     columns.append((u'CONDICIÓN PPL', 9000), )
#     columns.append((u'DISCAPACIDAD', 9000), )
#     # solo para encuesta con y sin discapacidad
#     if encuesta.id == 20 or encuesta.id == 21 or encuesta.id == 22 or encuesta.id == 23:  # encuestas
#         columns.append((u'TIENE DISCAPACIDAD', 3000), )
#     # fin
#
# if encuesta.tipoperfil == 2:  # DOCENTE
#     columns.append((u'Tipo de relación laboral', 9000), )
#     columns.append((u'Tiempo de dedicación', 9000), )
#     columns.append((u'Si es docente titular a qué categoría académica pertenece', 9000), )
#     # columns.append((u'MODALIDAD CONTRATACIÓN', 9000), )
#     columns.append((u'A qué Facultad pertenece', 9000), )
#     columns.append((u'Las carreras en las que imparte docencia actualmente ¿De qué modalidad son? ', 9000), )
#     # columns.append((u'MODALIDAD DE LA CARRERA QUE DESEARÍA TRABAJAR EN EL SEMESTRE 1S 2022', 9000), )
#
# if encuesta.tipoperfil == 3:  # ADMINISTRATIVO
#     columns.append((u'Tipo de relación laboral', 9000), )
#     columns.append((u'Denominación del puesto', 9000), )
#
# for x in preguntas:
#     columns.append((str(x.orden) + ") " + x.descripcion, 6000), )
#     if x.tipo == 1:
#         if not x.esta_vacia():
#             columns.append((str(x.orden) + ") " + x.observacionporno, 6000), )
# row_num = 3
# for col_num in range(len(columns)):
#     ws.write(row_num, col_num, columns[col_num][0], font_style)
#     ws.col(col_num).width = columns[col_num][1]
# row_num = 4
# i = 0
# datos=None
# if encuesta.tipoperfil == 1:  # ALUMNO
#     datos = encuesta.inscripcionencuestagrupoestudiantes_set.filter(status=True, respondio=True).order_by('inscripcion__persona__apellido1')
# if encuesta.tipoperfil == 2:  # DOCENTE
#     datos = encuesta.inscripcionencuestagrupoestudiantes_set.filter(status=True, respondio=True).order_by('profesor__persona__apellido1')
# if encuesta.tipoperfil == 3:  # ADMINISTRATIVO
#     datos = encuesta.inscripcionencuestagrupoestudiantes_set.filter(status=True, respondio=True).order_by('administrativo__persona__apellido1')
# cout_register = datos.count()
# register_start = 0
# limit = 0
# for dato in datos:
#     try:
#         row_num += limit
#         i += 1
#         limit = 0
#         ws.write(row_num, 0, i, font_style2)
#         if encuesta.tipoperfil == 1:
#             ws.write(row_num, 1, dato.inscripcion.persona.cedula, font_style2)
#             ws.write(row_num, 2, dato.inscripcion.persona.nombre_completo_inverso(), font_style2)
#         if encuesta.tipoperfil == 2:
#             ws.write(row_num, 1, dato.profesor.persona.cedula, font_style2)
#             ws.write(row_num, 2, dato.profesor.persona.nombre_completo_inverso(), font_style2)
#         if encuesta.tipoperfil == 3:
#             ws.write(row_num, 1, dato.administrativo.persona.cedula, font_style2)
#             ws.write(row_num, 2, dato.administrativo.persona.nombre_completo_inverso(), font_style2)
#         c = 3
#         if encuesta.tipoperfil == 1:
#             ws.write(row_num, c, dato.inscripcion.carrera.__str__(), font_style2) if not dato.inscripcion.carrera == None else ' '
#             c += 1
#             ws.write(row_num, c, MateriaAsignada.objects.filter(status=True, matricula__nivel__periodo_id=126,
#                                                                 matricula__inscripcion=dato.inscripcion).exclude(materia__asignaturamalla__malla_id=353).count(), font_style2)
#             c += 1
#             ws.write(row_num, c,u"SI" if dato.inscripcion.persona.ppl else "NO", font_style2)
#             c += 1
#             ws.write(row_num, c,u"SI" if dato.inscripcion.persona.tiene_discapasidad_new() else "NO", font_style2)
#             c += 1
#
#         # solo para encuesta con y sin discapacidad
#         if encuesta.id == 20 or encuesta.id == 21 or encuesta.id == 22 or encuesta.id == 23:  # encuesta
#             if dato.inscripcion.persona.mi_perfil().tienediscapacidad:
#                 discapacidad = 'SI'
#             else:
#                 discapacidad = 'NO'
#
#             ws.write(row_num, c, discapacidad, font_style2)
#             c += 1
#         # fin
#
#         if encuesta.tipoperfil == 2:
#             dt = ProfesorDistributivoHoras.objects.filter(status=True, periodo=126, profesor_id=dato.profesor.id).first()
#             ws.write(row_num, c, dt.nivelcategoria.nombre if dt is not None else '', font_style2)
#             c += 1
#             ws.write(row_num, c, dt.dedicacion.nombre if dt is not None else '', font_style2)
#             c += 1
#             ws.write(row_num, c, dt.categoria.nombre if dt is not None and dt.nivelcategoria.id == 1 else '', font_style2)
#             c += 1
#             ws.write(row_num, c, dt.coordinacion.nombre if dt is not None and dt.coordinacion is not None else '',
#                      font_style2)
#             c += 1
#             w = 0
#             for m in dato.profesor.mis_materias(126).values_list('materia__nivel__modalidad__nombre', flat=True).distinct(
#                     'materia__nivel__modalidad__nombre'):
#                 ws.write(row_num + w, c, str(m), font_style2)
#                 w += 1
#             if limit < w and w > 0:
#                 limit = w - 1
#
#             c += 1
#         if encuesta.tipoperfil == 3:
#             eDistributivoPersonas = DistributivoPersona.objects.filter(persona=dato.administrativo.persona, status=True,
#                                                                        regimenlaboral_id=2, estadopuesto_id=1)
#             eDistributivoPersona = None
#             if eDistributivoPersonas.values("id").exists():
#                 eDistributivoPersona = eDistributivoPersonas.first()
#             ws.write(row_num, c,
#                      eDistributivoPersona.regimenlaboral.descripcion if eDistributivoPersona is not None else '',
#                      font_style2)
#             c += 1
#             ws.write(row_num, c,
#                      eDistributivoPersona.denominacionpuesto.descripcion if eDistributivoPersona is not None else '',
#                      font_style2)
#             c += 1
#
#         for x in preguntas:
#             respuesta = None
#             if x.tipo == 1:
#                 respuesta = RespuestaPreguntaEncuestaGrupoEstudiantes.objects.filter(status=True, pregunta=x,
#                                                                                      inscripcionencuesta=dato)
#                 if respuesta.values("id").exists():
#                     respuesta = respuesta.first()
#                     ws.write(row_num, c, respuesta.respuesta, font_style2)
#                 else:
#                     respuesta = None
#                     ws.write(row_num, c, '', font_style2)
#                 c += 1
#
#             if x.tipo == 2:
#                 respuesta = dato.respuestarangoencuestagrupoestudiantes_set.get(status=True,
#                                                                                 pregunta=x) if dato.respuestarangoencuestagrupoestudiantes_set.filter(
#                     status=True, pregunta=x).exists() else None
#                 if respuesta is not None:
#                     rango = RangoPreguntaEncuestaGrupoEstudiantes.objects.get(pk=int(respuesta.opcionrango.id))
#                     ws.write(row_num, c, rango.descripcion, font_style2)
#                 else:
#                     ws.write(row_num, c, '', font_style2)
#                 c += 1
#             elif x.tipo == 5:
#                 respuesta = dato.respuestacuadriculaencuestagrupoestudiantes_set.filter(status=True,
#                                                                                         pregunta=x).first() if dato.respuestacuadriculaencuestagrupoestudiantes_set.filter(
#                     status=True, pregunta=x).exists() else None
#                 if respuesta is not None:
#                     try:
#                         int(respuesta.respuesta)
#                         if OpcionCuadriculaEncuestaGrupoEstudiantes.objects.filter(status=True, pregunta=x,
#                                                                                    id=respuesta.opcioncuadricula.id,
#                                                                                    tipoopcion=2).first() == None:
#                             resp = 'Sin contestar'
#                         else:
#                             resp = OpcionCuadriculaEncuestaGrupoEstudiantes.objects.filter(status=True, pregunta=x,
#                                                                                            id=respuesta.opcioncuadricula.id,
#                                                                                            tipoopcion=2).first().descripcion
#                     except ValueError:
#                         resp = respuesta.respuesta
#
#                     ws.write(row_num, c, resp, font_style2)
#                 else:
#                     ws.write(row_num, c, '', font_style2)
#                 c += 1
#             elif x.tipo == 6:
#                 respuesta = dato.respuestamultipleencuestagrupoestudiantes_set.filter(status=True,
#                                                                                       pregunta=x) if dato.respuestamultipleencuestagrupoestudiantes_set.values(
#                     'id').filter(status=True, pregunta=x).exists() else None
#                 if respuesta is not None:
#                     w = 0
#                     for rmult in respuesta:
#                         ws.write(row_num + w, c, rmult.opcionmultiple.descripcion, font_style2)
#                         # row_num += 1
#                         w += 1
#                     if limit < w and w > 0:
#                         limit = w - 1
#                 else:
#                     ws.write(row_num, c, '', font_style2)
#                 c += 1
#             else:
#                 if respuesta is not None:
#                     if x.tipo == 1:
#                         if not x.esta_vacia():
#                             ws.write(row_num, c, respuesta.respuestaporno if respuesta.respuestaporno else "", font_style2)
#                     else:
#                         ws.write(row_num, c, respuesta.respuesta, font_style2)
#                 else:
#                     ws.write(row_num, c, '', font_style2)
#                 c += 1
#         row_num += 1
#         print('%s' % (row_num))
#
#     except Exception as ex:
#         print('error: %s' % (ex))
#         pass
#
# wb.save(filename)
# print("FIN: ", filename)


# try:
#     miarchivo = openpyxl.load_workbook("lista ppl.xlsx")
#     lista = miarchivo.get_sheet_by_name('Hoja1')
#     totallista = lista.rows
#     a=0
#     for filas in totallista:
#         a += 1
#         if a > 2:
#             cedula = filas[0].value
#             crs=None
#             lider=None
#             correo_lider=None
#             telefono_lider=None
#             datospersona=None
#             if filas[5].value != None:
#                 crs = filas[5].value
#             if filas[6].value != None:
#                 lider = filas[6].value
#             if filas[7].value != None:
#                 correo_lider = filas[7].value
#             if filas[8].value != None:
#                 telefono_lider = filas[8].value
#             datospersona=None
#             if Persona.objects.filter(cedula=cedula).exists():
#                 datospersona = Persona.objects.get(cedula=cedula)
#             if Persona.objects.filter(pasaporte=cedula).exists():
#                 datospersona = Persona.objects.get(pasaporte=cedula)
#             if datospersona:
#                 if not datospersona.ppl:
#                     datospersona.ppl=True
#                     datospersona.save()
#                 if not HistorialPersonaPPL.objects.filter(status=True,persona=datospersona).exists():
#                     historial=HistorialPersonaPPL(persona=datospersona,centrorehabilitacion=crs,
#                                                   lidereducativo=lider,
#                                                   correolidereducativo=correo_lider,
#                                                   telefonolidereducativo=telefono_lider
#                     )
#                     historial.save()
#                 else:
#                     historial=HistorialPersonaPPL.objects.filter(status=True, persona=datospersona)[0]
#                     if not historial.centrorehabilitacion:
#                         historial.centrorehabilitacion = crs
#                     if not historial.lidereducativo:
#                         historial.lidereducativo = lider
#                     if not historial.correolidereducativo:
#                         historial.correolidereducativo = correo_lider
#                     if not historial.telefonolidereducativo:
#                         historial.telefonolidereducativo = telefono_lider
#                     historial.save()
#             print('Registro procesado: %s' %datospersona )
# except Exception as ex:
#         print('error: %s' % ex)


# import xlwt
# from xlwt import *
# from django.http import HttpResponse
#
# response = HttpResponse(content_type="application/ms-excel")
# response['Content-Disposition'] = 'attachment; filename=filename=reporte_horarios_examenes_sedes.xls'
# style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
# style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
# style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
# title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
# style1 = easyxf(num_format_str='D-MMM-YY')
# font_style = XFStyle()
# font_style.font.bold = True
# font_style2 = XFStyle()
# font_style2.font.bold = False
# wb = xlwt.Workbook()
# ws = wb.add_sheet('Sheetname')
# estilo = xlwt.easyxf('font: height 350, name Arial, colour_index black, bold on, italic on; align: wrap on, vert centre, horiz center;')
# ws.write_merge(0, 0, 0, 9, 'UNIVERSIDAD ESTATAL DE MILAGRO', estilo)
# output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
# nombre = "Lista" + datetime.now().strftime('%Y%m%d_%H%M%S') + ".xls"
# filename = os.path.join(output_folder, nombre)
# columns = [
#             (u"#", 1000),
#             (u"TIPO DOCUMENTO", 10000),
#             (u"DOCUMENTO", 10000),
#             (u"ALUMNO", 10000),
#             (u"CARRERA", 10000),
#             (u"NIVEL", 10000),
#             (u"ASIGNATURA", 10000),
#             (u"SEDE", 10000),
#             (u"FECHA", 6000),
#             (u"HORA INICIO", 6000),
#             (u"HORA FIN", 6000),
#             (u"SALA/LABORATORIO", 6000),
#             (u"RESPONSABLE", 6000),
#             (u"NUM. PLANIFICACIÓN", 6000),
# ]
# row_num = 3
# for col_num in range(len(columns)):
#     ws.write(row_num, col_num, columns[col_num][0], font_style)
#     ws.col(col_num).width = columns[col_num][1]
# row_num = 4
# ePeriodo = Periodo.objects.get(id=158)
# eMaterias = Materia.objects.filter(nivel__periodo=ePeriodo, asignaturamalla__malla__modalidad_id=3, status=True)
# eMateriaAsignadas = MateriaAsignada.objects.filter(materia__in=eMaterias)
# i = 0
# for eMateriaAsignada in eMateriaAsignadas:
#     i += 1
#     eMateriaAsignadaPlanificacionSedeVirtualExamenes = MateriaAsignadaPlanificacionSedeVirtualExamen.objects.filter(materiaasignada=eMateriaAsignada)
#     eMateria = eMateriaAsignada.materia
#     eAsignaturaMalla = eMateria.asignaturamalla
#     eNivelMalla = eAsignaturaMalla.nivelmalla
#     eAsignatura = eAsignaturaMalla.asignatura
#     eMatricula = eMateriaAsignada.matricula
#     eInscripcion = eMatricula.inscripcion
#     eCarrera = eInscripcion.carrera
#     ePersona = eInscripcion.persona
#     eLaboratorioVirtual = None
#     eResponsable = None
#     eTurnoPlanificacionSedeVirtualExamen = None
#     eFechaPlanificacionSedeVirtualExamen = None
#     eSedeVirtual = None
#     if eMateriaAsignadaPlanificacionSedeVirtualExamenes.values("id").exists():
#         eMateriaAsignadaPlanificacionSedeVirtualExamen = eMateriaAsignadaPlanificacionSedeVirtualExamenes.first()
#         eAulaPlanificacionSedeVirtualExamen = eMateriaAsignadaPlanificacionSedeVirtualExamen.aulaplanificacion
#         eLaboratorioVirtual = eAulaPlanificacionSedeVirtualExamen.aula
#         eResponsable = eAulaPlanificacionSedeVirtualExamen.responsable
#         eTurnoPlanificacionSedeVirtualExamen = eAulaPlanificacionSedeVirtualExamen.turnoplanificacion
#         eFechaPlanificacionSedeVirtualExamen = eTurnoPlanificacionSedeVirtualExamen.fechaplanificacion
#         eSedeVirtual = eFechaPlanificacionSedeVirtualExamen.sede
#     ws.write(row_num, 0, str(i), font_style2)
#     ws.write(row_num, 1, ePersona.tipo_documento(), font_style2)
#     ws.write(row_num, 2, ePersona.documento(), font_style2)
#     ws.write(row_num, 3, ePersona.nombre_completo(), font_style2)
#     ws.write(row_num, 4, eCarrera.nombrevisualizar if eCarrera.nombrevisualizar else eCarrera.nombre, font_style2)
#     ws.write(row_num, 5, eNivelMalla.nombre, font_style2)
#     ws.write(row_num, 6, eAsignatura.nombre, font_style2)
#     ws.write(row_num, 7, eSedeVirtual.nombre if eSedeVirtual else '', font_style2)
#     ws.write(row_num, 8, eFechaPlanificacionSedeVirtualExamen.fecha.__str__() if eFechaPlanificacionSedeVirtualExamen else '', font_style2)
#     ws.write(row_num, 9, eTurnoPlanificacionSedeVirtualExamen.horainicio.__str__() if eTurnoPlanificacionSedeVirtualExamen else '', font_style2)
#     ws.write(row_num, 10, eTurnoPlanificacionSedeVirtualExamen.horafin.__str__() if eTurnoPlanificacionSedeVirtualExamen else '', font_style2)
#     ws.write(row_num, 11, eLaboratorioVirtual.nombre if eLaboratorioVirtual else '', font_style2)
#     ws.write(row_num, 12, eResponsable.nombre_completo() if eResponsable else '', font_style2)
#     ws.write(row_num, 13, len(eMateriaAsignadaPlanificacionSedeVirtualExamenes.values("id")), font_style2)
#     row_num += 1
# wb.save(filename)
# print("FIN: ", filename)
# dataex = {}
# dataex['timeopen']=1661428800
# dataex['timeclose']=1661434199
# dataex['timelimit']=2100

# dataex['timeopen']=1661374800
# dataex['timeclose']=1661380200
# dataex['timelimit']=3600
# accesoquizindividual('aaguirrec6', Materia.objects.get(id=54586), 41625, dataex  )
# accesoquizindividual('crodriguezn', Materia.objects.get(id=54586), 41625, dataex  )
# password = random.randint(0, 999999999)
# print(u"%s"%password)
# accesoQuizIndividual('crodriguezn', Materia.objects.get(id=54586), 41625, dataex  )




# import xlwt
# from xlwt import *
# from django.http import HttpResponse
#
# response = HttpResponse(content_type="application/ms-excel")
# response['Content-Disposition'] = 'attachment; filename=silabo.xls'
# style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
# style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
# style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
# title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
# style1 = easyxf(num_format_str='D-MMM-YY')
# font_style = XFStyle()
# font_style.font.bold = True
# font_style2 = XFStyle()
# font_style2.font.bold = False
# wb = xlwt.Workbook()
# ws = wb.add_sheet('Sheetname')
# estilo = xlwt.easyxf('font: height 350, name Arial, colour_index black, bold on, italic on; align: wrap on, vert centre, horiz center;')
# ws.write_merge(0, 0, 0, 9, 'UNIVERSIDAD ESTATAL DE MILAGRO', estilo)
# output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
# nombre = "Lista" + datetime.now().strftime('%Y%m%d_%H%M%S') + ".xls"
# filename = os.path.join(output_folder, nombre)
# columns = [(u"FACULTAD", 6000),
#            (u"CARRERA", 6000),
#            (u"MATERIA", 6000),
#            (u"DOCENTE", 6000),
#            (u"TEMA SILABO", 6000),
#            (u"numsemana", 6000),
#            (u"fecha inicio semana", 6000),
#            (u"fecha fin semana", 6000),
#            (u"evaluacion_planificado", 6000),
#            (u"tema_actividad", 6000),
#            (u"foro_ingresado_aprobado", 6000),
#            (u"foro_moodle", 6000),
#            (u"tarea_ingresado_aprobado", 6000),
#            (u"tarea_moodle", 6000),
#            (u"test_ingresado_aprobado", 6000),
#            (u"test_moodle", 6000),
#            ]
# row_num = 3
# for col_num in range(len(columns)):
#     ws.write(row_num, col_num, columns[col_num][0], font_style)
#     ws.col(col_num).width = columns[col_num][1]
# row_num = 4
# cursor =  connections['sga_select'].cursor()
# sql="""
# SELECT co.nombre AS facultad, carr.nombre AS carrera, concat(asig.nombre,' ',mat.paralelo) AS materia,
# (
# SELECT concat(per.nombres,' ',per.apellido1,' ',per.apellido2) AS docente
# FROM sga_profesor prof
# INNER JOIN sga_persona per ON per.id=prof.persona_id
# WHERE prof.id=sil.profesor_id
# ) AS docente,
# (
# SELECT array_to_string(array_agg(tema.descripcion),' - ')
# FROM  sga_detallesilabosemanaltema dettema1
# INNER JOIN sga_temaunidadresultadoprogramaanalitico tema ON tema.id=dettema1.temaunidadresultadoprogramaanalitico_id
# WHERE dettema1."status" AND dettema1.silabosemanal_id=ss.id AND tema."status"
# ) AS tema_silabo,
# ss.numsemana,
# ss.fechainiciosemana,
# ss.fechafinciosemana,
# (
# SELECT array_to_string(array_agg(com.descripcion),',')
# FROM sga_evaluacionaprendizajesilabosemanal eval
# INNER JOIN sga_evaluacionaprendizajecomponente com ON com.id=eval.evaluacionaprendizaje_id
# WHERE eval.silabosemanal_id=ss.id AND eval."status"
# ) evaluacion_planificado,
# (
# SELECT array_to_string(array_agg(tema.descripcion),' - ')
# FROM sga_evaluacionaprendizajetema evaltema
# INNER JOIN sga_evaluacionaprendizajesilabosemanal eval1 ON eval1.id=evaltema.evaluacion_id
# INNER JOIN sga_detallesilabosemanaltema dettema ON dettema.id=evaltema.temasemanal_id
# INNER JOIN sga_temaunidadresultadoprogramaanalitico tema ON tema.id=dettema.temaunidadresultadoprogramaanalitico_id
# WHERE evaltema."status" AND eval1.silabosemanal_id=ss.id AND eval1."status" AND dettema.status AND tema.status
# ) AS tema_actividad,
# (
# SELECT array_to_string(array_agg(foro.nombre),',')
# FROM sga_forosilabosemanal foro
# WHERE foro.status AND foro.silabosemanal_id=ss.id AND foro.estado_id IN (1,2)
# ) AS foro_ingresado_aprobado,
# (
# SELECT array_to_string(array_agg(foro.nombre),',')
# FROM sga_forosilabosemanal foro
# WHERE foro.status AND foro.silabosemanal_id=ss.id AND foro.estado_id IN (4)
# ) AS foro_moodle,
# (
# SELECT array_to_string(array_agg(tarera.nombre),',')
# FROM sga_tareasilabosemanal tarera
# WHERE tarera.status AND tarera.silabosemanal_id=ss.id AND tarera.estado_id IN (1,2)
# ) AS tarea_ingresado_aprobado,
# (
# SELECT array_to_string(array_agg(tarea1.nombre),',')
# FROM sga_tareasilabosemanal tarea1
# WHERE tarea1.status AND tarea1.silabosemanal_id=ss.id AND tarea1.estado_id IN (4)
# ) AS tarea_moodle,
# (
# SELECT array_to_string(array_agg(test.nombretest),',')
# FROM sga_testsilabosemanal  test
# WHERE test.status AND test.silabosemanal_id=ss.id AND test.estado_id IN (1,2)
# ) AS test_ingresado_aprobado,
# (
# SELECT array_to_string(array_agg(test1.nombretest),',')
# FROM sga_testsilabosemanal test1
# WHERE test1.status AND test1.silabosemanal_id=ss.id and test1.estado_id IN (4)
# ) AS test_moodle
# FROM  sga_silabosemanal ss
# inner join sga_silabo sil ON sil.id=ss.silabo_id
# INNER JOIN sga_materia mat ON mat.id=sil.materia_id
# INNER JOIN sga_asignatura asig ON asig.id=mat.asignatura_id
# INNER JOIN sga_nivel niv ON niv.id=mat.nivel_id
# INNER JOIN sga_asignaturamalla asimall ON asimall.id=mat.asignaturamalla_id
# INNER JOIN sga_malla malla ON malla.id=asimall.malla_id
# INNER JOIN sga_carrera carr ON carr.id=malla.carrera_id
# INNER JOIN sga_coordinacion_carrera cc ON cc.carrera_id=carr.id
# INNER JOIN sga_coordinacion co ON co.id=cc.coordinacion_id
# WHERE sil."status" AND mat."status"  AND niv."status" AND niv.periodo_id=126 AND ss.status
# ORDER BY co.nombre,carr.nombre, asig.nombre,mat.paralelo,ss.numsemana
# """
# cursor.execute(sql)
# results = cursor.fetchall()
# for r in results:
#     try:
#         ws.write(row_num, 0, u'%s' % r[0], font_style2)
#         ws.write(row_num, 1, u'%s' % r[1], font_style2)
#         ws.write(row_num, 2, u'%s' % r[2], font_style2)
#         ws.write(row_num, 3, u'%s' % r[3], font_style2)
#         ws.write(row_num, 4, u'%s' % r[4], font_style2)
#         ws.write(row_num, 5, u'%s' % r[5], font_style2)
#         ws.write(row_num, 6, u'%s' % r[6], font_style2)
#         ws.write(row_num, 7, u'%s' % r[7], font_style2)
#         ws.write(row_num, 8, u'%s' % r[8], font_style2)
#         ws.write(row_num, 9, u'%s' % r[9], font_style2)
#         ws.write(row_num, 10, u'%s' % r[10], font_style2)
#         ws.write(row_num, 11, u'%s' % r[11], font_style2)
#         ws.write(row_num, 12, u'%s' % r[12], font_style2)
#         ws.write(row_num, 13, u'%s' % r[13], font_style2)
#         ws.write(row_num, 14, u'%s' % r[14], font_style2)
#         ws.write(row_num, 15, u'%s' % r[15], font_style2)
#         row_num+=1
#         print('%s' % (row_num))
#     except Exception as ex:
#         print('error: %s' % (ex))
#         pass
# wb.save(filename)
# print("FIN: ", filename)
# matriculas=MateriaAsignada.objects.values_list('matricula_id',flat=True).filter(id__in=[2296962,2292995])
# for matricula in Matricula.objects.filter(id__in=matriculas):
#     print(u"*** Matricula (ID: %s - %s)" % (matricula.id, matricula))
#     matricula.actualiza_matricula()
#     print(u"*** Estado de matricula actualizado Matricula (ID: %s)" % matricula.id)
#
#     if matricula.bloqueomatricula and not matricula.retiradomatricula:
#         rubrospagados = matricula.rubro_set.values('id').filter(cancelado=True, status=True).count()
#         rubrosdebe = matricula.rubro_set.values('id').filter(status=True).count()
#         if rubrospagados == rubrosdebe:
#             matricula.bloqueomatricula = False
#             matricula.save()
#             usermoodle = matricula.inscripcion.persona.usuario.username
#             cnmoodle = connections['moodle_db'].cursor()
#             if usermoodle and cnmoodle:
#                 # Consulta en mooc_user
#                 sql = """Select id From mooc_user Where suspended=1 and username='%s'""" % (usermoodle)
#                 cnmoodle.execute(sql)
#                 registro = cnmoodle.fetchall()
#                 if registro:
#                     sql = """Update mooc_user Set suspended=0 Where username='%s'""" % (usermoodle)
#                     cnmoodle.execute(sql)

# import xlwt
# from xlwt import *
# from django.http import HttpResponse
#
# response = HttpResponse(content_type="application/ms-excel")
# response['Content-Disposition'] = 'attachment; filename=reporte_notas_ingles.xls'
# style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
# style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
# style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
# title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
# style1 = easyxf(num_format_str='D-MMM-YY')
# font_style = XFStyle()
# font_style.font.bold = True
# font_style2 = XFStyle()
# font_style2.font.bold = False
# wb = xlwt.Workbook()
# ws = wb.add_sheet('Sheetname')
# estilo = xlwt.easyxf('font: height 350, name Arial, colour_index black, bold on, italic on; align: wrap on, vert centre, horiz center;')
# ws.write_merge(0, 0, 0, 9, 'UNIVERSIDAD ESTATAL DE MILAGRO', estilo)
# output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
# nombre = "Lista" + datetime.now().strftime('%Y%m%d_%H%M%S') + ".xls"
# filename = os.path.join(output_folder, nombre)
# columns = [(u"CEDULA", 6000),
#            (u"APELLIDOS Y NOMBRES", 6000),
#            (u"CARRERA", 6000),
#            (u"NOTA", 6000),
#            (u"ESTADO", 6000)
#            ]
# row_num = 3
# for col_num in range(len(columns)):
#     ws.write(row_num, col_num, columns[col_num][0], font_style)
#     ws.col(col_num).width = columns[col_num][1]
# row_num = 4
# cont=0
# for materiaasignada in MateriaAsignada.objects.filter(status=True, materia__nivel_id=719, materia__status=True,retiramateria=False):
#     cont+=1
#     url = 'https://upei.buckcenter.edu.ec/usernametograde.php?username=%s'%(materiaasignada.matricula.inscripcion.persona.identificacion())
#     ws.write(row_num, 0, materiaasignada.matricula.inscripcion.persona.cedula)
#     ws.write(row_num, 1, materiaasignada.matricula.inscripcion.persona.apellido1 + ' ' + materiaasignada.matricula.inscripcion.persona.apellido2 + ' ' + materiaasignada.matricula.inscripcion.persona.nombres)
#     ws.write(row_num, 2, str(materiaasignada.matricula.inscripcion.carrera))
#     try:
#             print("FILA %s: " % cont)
#             req = Request(url)
#             response = urlopen(req)
#             result = json.loads(response.read().decode())
#             idcurso = int(result['idcurso'])
#             nota=0
#             try:
#                 nota = float(result['notam'])
#             except:
#                 nota=0
#             print(u"%s" % result)
#             print(u"ID CURSO: %s" % idcurso)
#             print(u"NOTA: %s" % nota)
#             ws.write(row_num, 3, u"%s" %nota)
#             ws.write(row_num, 4, u"APROBADO" if nota>=70 else "REPROBADO")
#             ws.write(row_num, 5, u"COINCIDE CURSO" if idcurso==materiaasignada.materia.idcursomoodle else "NO COINCIDE CURSO")
#     except Exception as ex:
#         print('error: %s' % (ex))
#         ws.write(row_num, 3, u"%s" % ex)
#         ws.write(row_num, 4, u"%s" % ex)
#         ws.write(row_num, 5, u"%s" % ex)
#         pass
#     row_num += 1
# wb.save(filename)
# print("FIN: ", filename)


def get_valor_cobro_matricula(eMalla):
    from matricula.models import CostoOptimoMalla
    ePeriodo = Periodo.objects.get(id=153)
    costomatricula=0
    eCostoOptimoMallas = CostoOptimoMalla.objects.filter(status=True, periodo=ePeriodo, malla=eMalla)
    if eCostoOptimoMallas.values("id").exists():
        costomatricula = eCostoOptimoMallas[0].costomatricula
    if costomatricula == 0 or costomatricula is None:
        return 0
    return costomatricula

def get_valor_crobro_arancel_nse(eMalla,nivelmalla,egruposocioeconomico):
    from matricula.models import CostoOptimoNivelMalla,CostoOptimoGrupoSocioEconomico
    valor_arancel=0
    eCostoOptimoNivelMallas = CostoOptimoNivelMalla.objects.filter(status=True,
                                                                   nivelmalla=nivelmalla,
                                                                   costooptimomalla__periodo=153,
                                                                   costooptimomalla__malla=eMalla)
    if eCostoOptimoNivelMallas.values("id").exists():
        ecostooptimonivelmalla = eCostoOptimoNivelMallas[0]
        cogse=CostoOptimoGrupoSocioEconomico.objects.filter(status=True, costooptimonivelmalla=ecostooptimonivelmalla,
                                                      gruposocioeconomico=egruposocioeconomico)
        if cogse.exists():
            valor_arancel=cogse[0].costoarancel
    return valor_arancel


# ESTIMADO DE COBROS

import xlwt
from xlwt import *
def reporte_estimado_cobros_matricula(periodo):
    from django.http import HttpResponse
    from sga.funciones import null_to_decimal
    response = HttpResponse(content_type="application/ms-excel")
    response['Content-Disposition'] = 'attachment; filename=estimado_cobro.xls'
    style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
    style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
    style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
    title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
    style1 = easyxf(num_format_str='D-MMM-YY')
    font_style = XFStyle()
    font_style.font.bold = True
    font_style2 = XFStyle()
    font_style2.font.bold = False
    wb = xlwt.Workbook()
    ws = wb.add_sheet('Sheetname')
    estilo = xlwt.easyxf('font: height 350, name Arial, colour_index black, bold on, italic on; align: wrap on, vert centre, horiz center;')
    ws.write_merge(0, 0, 0, 9, 'UNIVERSIDAD ESTATAL DE MILAGRO', estilo)
    output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
    nombre = "Lista" + datetime.now().strftime('%Y%m%d_%H%M%S') + ".xls"
    filename = os.path.join(output_folder, nombre)
    columns = [(u"CEDULA", 6000),
               (u"APELLIDOS Y NOMBRES", 6000),
               (u"CARRERA", 6000),
               (u"ASIGNATURA", 6000),
               (u"CRÉDITOS", 6000),
               (u"HORAS", 6000),
               (u"NIVEL", 6000),
               (u"MATERIA", 6000),
               (u"GRUPO SOCIO ECONÓMICO", 6000),
               (u"COSTO MATRÍCULA", 6000),
               (u"VALOR GRUPO SOCIO ECONÓMICO", 6000),
               (u"TOTAL", 6000),
               (u"TIPO", 6000)
               ]
    row_num = 3
    for col_num in range(len(columns)):
        ws.write(row_num, col_num, columns[col_num][0], font_style)
        ws.col(col_num).width = columns[col_num][1]
    row_num = 4
    cont=0
    matriculas_perdida=Matricula.objects.filter(status=True, nivel__periodo=periodo,retiradomatricula=False)
    lista_matricula=[]
    for matricula in matriculas_perdida:
        if not matricula.id in lista_matricula:
            matriculas_periodo=Matricula.objects.filter(status=True, nivel__periodo=periodo, inscripcion__persona=matricula.inscripcion.persona,retiradomatricula=False)
            cantidad_matricula=matriculas_periodo.count()
            if matricula.inscripcion.estado_gratuidad==3 or cantidad_matricula>1:
                if cantidad_matricula>1:
                    matricula = matriculas_periodo.order_by('-inscripcion__fecha')[0]
                    for mat2 in matriculas_periodo:
                        lista_matricula.append(mat2.id)

                else:
                    lista_matricula.append(matricula.id)
                eMalla = matricula.inscripcion.mi_malla()
                gse = matricula.matriculagruposocioeconomico()
                materiaasignadas=MateriaAsignada.objects.filter(status=True,matricula=matricula,retiramateria=False)
                for materiaasignada in materiaasignadas:
                    try:
                        cont += 1
                        asignaturamalla=materiaasignada.materia.asignaturamalla
                        costo_arancel_nse = get_valor_crobro_arancel_nse(eMalla,asignaturamalla.nivelmalla,gse)
                        ws.write(row_num, 0, materiaasignada.matricula.inscripcion.persona.cedula)
                        ws.write(row_num, 1, materiaasignada.matricula.inscripcion.persona.nombre_completo())
                        ws.write(row_num, 2, str(materiaasignada.matricula.inscripcion.carrera))
                        ws.write(row_num, 3, str(asignaturamalla.asignatura))
                        ws.write(row_num, 4, str(asignaturamalla.creditos))
                        ws.write(row_num, 5, str(asignaturamalla.horas))
                        ws.write(row_num, 6, str(asignaturamalla.nivelmalla))
                        ws.write(row_num, 7, str(materiaasignada.materia))
                        ws.write(row_num, 8, str(materiaasignada.matricula.matriculagruposocioeconomico()))
                        ws.write(row_num, 9, str(get_valor_cobro_matricula(eMalla)))
                        ws.write(row_num, 10, str(costo_arancel_nse))
                        ws.write(row_num, 11, str(null_to_decimal(float(costo_arancel_nse)*float(asignaturamalla.creditos),2)))
                        ws.write(row_num, 12, str("PERDIDA GRATUIDAD"))
                        row_num+=1
                        print(u"%s"%cont)
                    except Exception as ex:
                        print('error: %s' % ex)

    # cont=0
    # materiaasignadas=MateriaAsignada.objects.filter(status=True,
    #                                                 materia__nivel__periodo=periodo,
    #                                                 estado_id=2,
    #                                                 matriculas__lte=3).order_by('matricula')
    # print(u"%s"%materiaasignadas.count())
    # for materiaasignada in materiaasignadas:
    #     try:
    #         cont += 1
    #         if not materiaasignada.matricula.inscripcion.estado_gratuidad == 3:
    #             lista_matricula.append(materiaasignada.matricula.id)
    #             eMalla = materiaasignada.matricula.inscripcion.mi_malla()
    #             asignaturamalla=materiaasignada.materia.asignaturamalla
    #             gse=materiaasignada.matricula.matriculagruposocioeconomico()
    #             costo_arancel_nse = get_valor_crobro_arancel_nse(eMalla,asignaturamalla.nivelmalla,gse)
    #             ws.write(row_num, 0, materiaasignada.matricula.inscripcion.persona.cedula)
    #             ws.write(row_num, 1, materiaasignada.matricula.inscripcion.persona.nombre_completo())
    #             ws.write(row_num, 2, str(materiaasignada.matricula.inscripcion.carrera))
    #             ws.write(row_num, 3, str(asignaturamalla.asignatura))
    #             ws.write(row_num, 4, str(asignaturamalla.creditos))
    #             ws.write(row_num, 5, str(asignaturamalla.horas))
    #             ws.write(row_num, 6, str(asignaturamalla.nivelmalla))
    #             ws.write(row_num, 7, str(materiaasignada.materia))
    #             ws.write(row_num, 8, str(materiaasignada.matricula.matriculagruposocioeconomico()))
    #             ws.write(row_num, 9, str(get_valor_cobro_matricula(eMalla)))
    #             ws.write(row_num, 10, str(costo_arancel_nse))
    #             ws.write(row_num, 11, str(null_to_decimal(float(costo_arancel_nse)*float(asignaturamalla.creditos),2)))
    #             ws.write(row_num, 12, str("REPROBADO"))
    #             row_num+=1
    #             print(u"%s"%cont)
    #     except Exception as ex:
    #         print('error: %s' % ex)

    wb.save(filename)
    print("FIN: ", filename)

# periodo=Periodo.objects.get(id=153)
# reporte_estimado_cobros_matricula(periodo)


# arra = [
# {'idr': 2932059, 'nota': 86, 'asistencia': 80, 'estado': 'APROBADO'},
# {'idr': 2919062, 'nota': 88, 'asistencia': 93, 'estado': 'APROBADO'},
# {'idr': 2939486, 'nota': 92, 'asistencia': 100, 'estado': 'APROBADO'},
# {'idr': 2940444, 'nota': 77, 'asistencia': 95, 'estado': 'APROBADO'},
# {'idr': 2941498, 'nota': 77, 'asistencia': 95, 'estado': 'APROBADO'},
# {'idr': 2938462, 'nota': 83, 'asistencia': 91, 'estado': 'APROBADO'},
# {'idr': 2930740, 'nota': 96, 'asistencia': 100, 'estado': 'APROBADO'},
# {'idr': 2935872, 'nota': 84, 'asistencia': 88, 'estado': 'APROBADO'},
# {'idr': 2921637, 'nota': 91, 'asistencia': 100, 'estado': 'APROBADO'},
# {'idr': 2940979, 'nota': 84, 'asistencia': 100, 'estado': 'APROBADO'},
# {'idr': 2926557, 'nota': 80, 'asistencia': 92, 'estado': 'APROBADO'},
# {'idr': 2916238, 'nota': 91, 'asistencia': 100, 'estado': 'APROBADO'},
# {'idr': 2917115, 'nota': 91, 'asistencia': 100, 'estado': 'APROBADO'},
# {'idr': 2917195, 'nota': 88, 'asistencia': 92, 'estado': 'APROBADO'},
# {'idr': 2917751, 'nota': 95, 'asistencia': 100, 'estado': 'APROBADO'},
# {'idr': 2919916, 'nota': 89, 'asistencia': 100, 'estado': 'APROBADO'},
# {'idr': 2920993, 'nota': 81, 'asistencia': 92, 'estado': 'APROBADO'},
# {'idr': 2924694, 'nota': 96, 'asistencia': 100, 'estado': 'APROBADO'},
# {'idr': 2925052, 'nota': 93, 'asistencia': 100, 'estado': 'APROBADO'},
# {'idr': 2926534, 'nota': 92, 'asistencia': 100, 'estado': 'APROBADO'},
# {'idr': 2931494, 'nota': 80, 'asistencia': 92, 'estado': 'APROBADO'},
# {'idr': 2937532, 'nota': 81, 'asistencia': 91, 'estado': 'APROBADO'},
# {'idr': 2934857, 'nota': 92, 'asistencia': 94, 'estado': 'APROBADO'},
# {'idr': 2943830, 'nota': 78, 'asistencia': 96, 'estado': 'APROBADO'},
# {'idr': 2941557, 'nota': 93, 'asistencia': 100, 'estado': 'APROBADO'},
# {'idr': 2944428, 'nota': 85, 'asistencia': 95, 'estado': 'APROBADO'},
# {'idr': 2944654, 'nota': 95, 'asistencia': 100, 'estado': 'APROBADO'},
# {'idr': 2944145, 'nota': 95, 'asistencia': 97, 'estado': 'APROBADO'},
# {'idr': 2915834, 'nota': 92, 'asistencia': 92, 'estado': 'APROBADO'},
# {'idr': 2917130, 'nota': 74, 'asistencia': 100, 'estado': 'APROBADO'},
# {'idr': 2921264, 'nota': 54, 'asistencia': 96, 'estado': 'REPROBADO '},
# {'idr': 2923105, 'nota': 87, 'asistencia': 100, 'estado': 'APROBADO'},
# {'idr': 2926272, 'nota': 57, 'asistencia': 100, 'estado': 'REPROBADO '},
# {'idr': 2926819, 'nota': 78, 'asistencia': 90, 'estado': 'APROBADO'},
# {'idr': 2928209, 'nota': 71, 'asistencia': 98, 'estado': 'APROBADO'},
# {'idr': 2930166, 'nota': 50, 'asistencia': 96, 'estado': 'REPROBADO '},
# {'idr': 2934155, 'nota': 92, 'asistencia': 100, 'estado': 'APROBADO'},
# {'idr': 2916148, 'nota': 80, 'asistencia': 94, 'estado': 'APROBADO'},
# {'idr': 2916597, 'nota': 93, 'asistencia': 100, 'estado': 'APROBADO'},
# {'idr': 2917933, 'nota': 94, 'asistencia': 100, 'estado': 'APROBADO'},
# {'idr': 2919710, 'nota': 96, 'asistencia': 92, 'estado': 'APROBADO'},
# {'idr': 2919992, 'nota': 92, 'asistencia': 94, 'estado': 'APROBADO'},
# {'idr': 2921107, 'nota': 90, 'asistencia': 100, 'estado': 'APROBADO'},
# {'idr': 2922003, 'nota': 77, 'asistencia': 86, 'estado': 'APROBADO'},
# {'idr': 2928455, 'nota': 90, 'asistencia': 90, 'estado': 'APROBADO'},
# {'idr': 2930619, 'nota': 82, 'asistencia': 97, 'estado': 'APROBADO'},
# {'idr': 2930693, 'nota': 83, 'asistencia': 98, 'estado': 'APROBADO'},
# {'idr': 2932022, 'nota': 88, 'asistencia': 100, 'estado': 'APROBADO'},
# {'idr': 2935473, 'nota': 85, 'asistencia': 92, 'estado': 'APROBADO'},
# {'idr': 2940831, 'nota': 35, 'asistencia': 100, 'estado': 'REPROBADO '},
# {'idr': 2919636, 'nota': 72, 'asistencia': 100, 'estado': 'APROBADO'},
# {'idr': 2925386, 'nota': 78, 'asistencia': 91, 'estado': 'APROBADO'},
# {'idr': 2926399, 'nota': 52, 'asistencia': 100, 'estado': 'REPROBADO '},
# {'idr': 2927676, 'nota': 73, 'asistencia': 86, 'estado': 'APROBADO'}
# ]

# for x in arra:
#     record = RecordAcademico.objects.get(id=x.get("idr"))
#     record.inscripcion.actualizar_creditos()
#     record.inscripcion.actualizar_nivel()
# print(u"fin")

# CREACIÓN DE CURSOS MATERIAS INGLÉS
# malla=Malla.objects.get(id=353)
# nivel=Nivel.objects.get(id=764)
# for asignaturamalla in AsignaturaMalla.objects.filter(status=True,malla=malla,asignatura_id=1691):
#     for cont in range(40):
#         paralelos=Paralelo.objects.filter(status=True,nombre='M6.G%s'%(cont+1))
#         if paralelos.exists():
#             paralelo=paralelos[0]
#             if not Materia.objects.filter(status=True,nivel=nivel, asignatura=asignaturamalla.asignatura, asignaturamalla=asignaturamalla,paralelomateria=paralelo).exists():
#                 materia = Materia(asignatura=asignaturamalla.asignatura,
#                                   asignaturamalla=asignaturamalla,
#                                   nivel=nivel,
#                                   horas=asignaturamalla.horas,
#                                   horassemanales=0,
#                                   creditos=asignaturamalla.creditos,
#                                   identificacion=asignaturamalla.identificacion,
#                                   alias=u"INGLÉS 06",
#                                   paralelo=paralelo.nombre,
#                                   paralelomateria=paralelo,
#                                   parcial=0,
#                                   inicio=datetime(2022, 11, 3, 0, 0, 0).date(),
#                                   fin=datetime(2022, 12, 8, 0, 0, 0).date(),
#                                   cerrado=False,
#                                   actualizarhtml=False,
#                                   rectora=False,
#                                   practicas=False,
#                                   tutoria=False,
#                                   grado=False,
#                                   validacreditos=True,
#                                   validapromedio=True,
#                                   tipomateria=2,
#                                   seevalua=False,
#                                   modeloevaluativo_id=23,
#                                   cupo=200,
#                                   esintroductoria=False,
#                                   inglesepunemi=True,
#                                   )
#                 materia.save()
#                 print(u"%s"%materia)

# detalle de valores
# nombre_periodo=Periodo.objects.get(id=153).nombre
# periodo = Periodo.objects.get(id=126)
# for matricula in Matricula.objects.filter(status=True, nivel__periodo_id=153,retiradomatricula=False).exclude(inscripcion__carrera__id__in=[7,138,129,90,157]):
#     rubros= Rubro.objects.filter(status=True,matricula=matricula,observacion="INGLÉS %s" %(nombre_periodo))
#     if rubros.exists():
#         rubro=rubros[0]
#         materiaasigandas=rubro.materiaasignada_set.filter(status=True,materia__inglesepunemi=True,materia__nivel_id=764)
#         if materiaasigandas.exists():
#             materiaasignada=materiaasigandas[0]
#             malla = matricula.inscripcion.mi_malla()
#             if malla:
#                 periodomalla = PeriodoMalla.objects.filter(periodo=periodo, malla=malla, status=True)
#                 if periodomalla.values("id").exists():
#                     periodomalla = periodomalla[0]
#                     matriculagruposocioeconomico = matricula.matriculagruposocioeconomico_set.filter(status=True)
#
#                     if matriculagruposocioeconomico.values("id").exists():
#                         gruposocioeconomico = matriculagruposocioeconomico[0].gruposocioeconomico
#                     else:
#                         gruposocioeconomico = matricula.inscripcion.persona.grupoeconomico()
#
#                     detalleperiodomalla = DetallePeriodoMalla.objects.filter(periodomalla=periodomalla,
#                                                                              gruposocioeconomico=gruposocioeconomico,
#                                                                              status=True)
#                     if detalleperiodomalla.values("id").exists():
#                         valorgrupoeconomico = detalleperiodomalla[0].valor
#                         creditos_para_cobro = materiaasignada.materia.creditos
#                         if materiaasignada.existe_modulo_en_malla():
#                             creditos_para_cobro = materiaasignada.materia_modulo_malla().creditos
#                             if not DetalleRubroMatricula.objects.filter(status=True,matricula=matricula,materia=materiaasignada.materia).exists():
#                                 detallerubros=DetalleRubroMatricula(matricula=matricula,
#                                                       materia=materiaasignada.materia,
#                                                       costo=valorgrupoeconomico,
#                                                       creditos=creditos_para_cobro)
#                                 detallerubros.save()

# materiasasigandas=MateriaAsignada.objects.filter(status=True,materia__inglesepunemi=True,materia__nivel_id=764,materia__asignatura_id=1691)
# print(materiasasigandas.count())
# cont=0
# for materiaasignada in materiasasigandas:
#     cont+=1
#     materiaasignada.delete()
#     print(cont)


# matricula=Matricula.objects.get(id=524197)
# materiaasignada=MateriaAsignada.objects.get(id=2639968)
# matricula.nuevo_calculo_matricula_ingles(materiaasignada)


# nombre_periodo=Periodo.objects.get(id=153).nombre
# for matricula in Matricula.objects.filter(status=True, nivel__periodo_id=153,retiradomatricula=False,id__in=[535189,
# 547474,
# 552249,
# 530347]):
#     materiasasignadas= MateriaAsignada.objects.filter(status=True,matricula=matricula,materia__asignatura_id=1691,materia__inglesepunemi=True,materia__nivel_id=764)
#     if not materiasasignadas.exists():
#         rubros= Rubro.objects.filter(status=True,matricula=matricula,observacion="INGLÉS %s" %(nombre_periodo))
#         if rubros.exists():
#             rubros.delete()
#             print(u"%s"%matricula)



# try:
#     miarchivo = openpyxl.load_workbook("MTN Final de cupos 2s-2022.xlsx")
#     lista = miarchivo.get_sheet_by_name('Hoja1')
#     totallista = lista.rows
#     a=0
#     for filas in totallista:
#         a += 1
#         if a > 2:
#             cedula = str(filas[14].value)
#             genero = int(filas[20].value)
#             if Persona.objects.filter(cedula=cedula).exists():
#                 datospersona = Persona.objects.get(cedula=cedula)
#             if Persona.objects.filter(pasaporte=cedula).exists():
#                 datospersona = Persona.objects.get(pasaporte=cedula)
#             if datospersona:
#                 datospersona.sexo_id=genero
#                 datospersona.save()
#                 print('Registro actualizado: %s' %datospersona )
# except Exception as ex:
#         print('error: %s' % ex)
# materias=Materia.objects.filter(status=True,nivel__periodo_id=153,asignaturamalla__malla__carrera__id__in=[110],idcursomoodle__gt=0).order_by('asignaturamalla__malla__carrera','asignaturamalla__nivelmalla')
# print(u"%s"%materias.count())
# for materia in materias:
#     materia.crear_actualizar_estudiantes_curso(moodle, 1)

# fecha=datetime(2022, 11,17, 0, 0, 0).date()
# matriculas=Matricula.objects.filter(status=True, nivel__periodo_id=202,fecha_creacion__gte=fecha)
# print(matriculas.count())
# contador=0
# for matricula in matriculas:
#     contador+=1
#     print("%s"%contador)
#     print(u"%s" % matricula)
#     matricula.delete()
# ria.objects.get(id=60270)
# tiot=materia.materiaasignada_set.filter(retiramateria=False,matricula__status=True)
# print(tiot.count())
# materia=Mate

# periodomatricula = PeriodoMatricula.objects.filter(status=True, periodo_id=153)
# periodomatricula = periodomatricula[0]
# configuracion_carnet = periodomatricula.configuracion_carnet
# matriculas=Matricula.objects.filter(status=True,inscripcion__carrera_id=110, nivel__periodo_id=153)
# print(matriculas.count())
# contador=0
# for matricula in matriculas:
#     contador+=1
#     print("%s - %s" %(contador, matricula))
#     carne = Carnet.objects.filter(config=configuracion_carnet, persona=matricula.inscripcion.persona, matricula=matricula)
#     if carne.exists():
#         carne.delete()


# matriculas=Matricula.objects.filter(status=True,termino=False,nivel__periodo_id=202)
# print(matriculas.count())
# cont=0
# for matricula in matriculas:
#     cont+=1
#     print("%s -%s"%(cont,matricula))
#     matricula.delete()

# materiasasignadas=MateriaAsignada.objects.filter(id__in=[2614688,
# 2615058,
# 2627800])
# for materiaasgnada in materiasasignadas:
#     matricula=materiaasgnada.matricula
#     if DetalleRubroMatricula.objects.filter(status=True,matricula=matricula,materia=materiaasgnada.materia).exists():
#         detalle=DetalleRubroMatricula.objects.get(status=True,matricula=matricula,materia=materiaasgnada.materia)
#         detalle.activo=True
#         detalle.save()
#         # matricula.actualiza_matricula()
#         print("%s"%materiaasgnada)

# materias = Materia.objects.filter(id__in=[61328,
# 61463])
# print(u"%s"%materias.count())
# for materia in materias:
#     print(materia)
#     materia.delete()

# materiaasignada=MateriaAsignada.objects.get(id=2613411)
# materiaasignada.delete()
# matricula=materiaasignada.matricula
# if DetalleRubroMatricula.objects.filter(status=True,matricula=matricula,materia=materiaasignada.materia).exists():
#         detalle=DetalleRubroMatricula.objects.get(status=True,matricula=matricula,materia=materiaasignada.materia)
#         detalle.activo=False
#         detalle.save()
#         matricula.actualiza_matricula()

# tipourl = 2
# periodo = Periodo.objects.get(pk=202)
# cursos = Materia.objects.filter(nivel__periodo=periodo, status=True, materiaasignada__status=True,asignaturamalla__asignatura__id=4837).order_by("idcursomoodle", 'asignatura__nombre', 'inicio', 'identificacion', 'id').distinct()
# for curso in cursos:
#         curso.crear_actualizar_estudiantes_curso_admision(moodle, tipourl)
# 60949
# materias=Materia.objects.filter(id=60950)
# for materiaasignada in MateriaAsignada.objects.filter(materia__in=materias):
#     rubro=Rubro.objects.filter(matricula=materiaasignada.matricula,observacion = 'INGLÉS NOVIEMBRE 2022 MARZO 2023')
#     if rubro.exists():
#         print(materiaasignada)
    # historico=HistoricoRecordAcademico.objects.filter(inscripcion=materiaasignada.matricula.inscripcion,
    #                                         materiaregular=materiaasignada.materia,
    #                                         fecha__gte=datetime(2022,12,12,0,0,0).date())
    #
    #
    # if historico:
    #     record=historico[0].recordacademico
    #     historico.delete()
    #     record.actualizar()
# import csv
# from time import sleep
# def importar_data():
#     VARIABLES = (
#         (0, u"PERIODO"),
#         (1, u"IDENTIFICACION"), #persona
#         (2, u"TIPO_DOCUMENTO"), #persona
#         (3, u"NOMBRES"), #persona
#         (4, u"APELLIDOS"), #persona
#         (5, u"NACIONALIDAD"), #persona
#         (6, u"FECHA_NACIMIENTO"), #persona
#         (7, u"ESTADO_CIVIL"), #persona
#         (8, u"ESTADO_REGISTRO_NACIONAL"), #aplicante
#         (9, u"FECHA_FIN_REGISTRO_NACIONAL"), #aplicante
#         (10, u"SEXO"), #persona
#         (11, u"GENERO"), #persona
#         (12, u"CODIGO_PROVINCIA_RESIDE"),
#         (13, u"CODIGO_CANTON_RESIDE"),
#         (14, u"CANTON_RESIDE"), #persona
#         (15, u"CODIGO_PARROQUIA_RESIDE"),
#         (16, u"PARROQUIA_RESIDE"), #persona
#         (17, u"PAIS_DESEA_RENDIR"), #persona
#         (19, u"NOMBRE_CONSULADO"), #persona
#         (20, u"AUTOIDENTIFICACION"), #personaetnia
#         (21, u"PUEBLO_INDIGENA"),#personaetnia
#         (22, u"TIENE_INTERNET_EN_DOMICILIO"),#aplicante
#         (23, u"COMPUTADORA_EN_DOMICILIO"),#aplicante
#         (24, u"SISTEMA_OPERATIVO"),#aplicante
#         (25, u"CAMARA_WEB"),#aplicante
#         (26, u"REQUIERE_ASIGNE_COMPUTADORA"),#aplicante
#         (27, u"CARNET_CONADIS"),#persona discapacidad
#         (28, u"TIPO_DISCAPACIDAD"),#persona discapacidad
#         (29, u"PORCENTAJE_DISCAPACIDAD"),#persona discapacidad
#         (30, u"REQUIERE_APOYO_DISCAPACIDAD"),#persona discapacidad
#         (31, u"CONSTA_REGISTRO_SOCIAL"), #aplicante SI O NO BOOLEAN
#         (32, u"NOMBRE_IES"),#aplicante
#         (33, u"NOMBRE_CAMPO_AMPLIO"),#aplicante
#         (34, u"PPL"),#PERSONA PPL
#         (35, u"NOMBRE_CENTRO_PPL_CAI")#PERSONA PPL
#     )
#     cont = 0
#     with open('registroNacionalP25.csv', newline='', encoding="utf8") as File:
#         reader = csv.reader(File)
#         for row in reader:
#             if cont>0:
#                 print(str(row))
#                 texto=str(row[0])
#                 texto=texto.split('|')
#                 for x in range(35):
#                     try:
#                         print(texto[x])
#                     except Exception as ex:
#                         print("ERROR" + str(ex))
#                         time.sleep(15)
#             cont=+1
#
# importar_data()
# from datetime import datetime
# import time
# conexion = connections['moodle_db']
# cursor = conexion.cursor()
# inicio = datetime(2023, 3, 3, 15, 0, 3)
# inicio = int(time.mktime(inicio.timetuple()))
# fin = datetime(2023, 3, 3, 15, 18, 20)
# fin = int(time.mktime(fin.timetuple()))
# sql = """update mooc_grade_grades set timecreated=%s, timemodified=%s where id=1836109""" % (inicio, fin)
# print(sql)
# cursor.execute(sql)
# fecha = datetime.now()
# fecha = int(time.mktime(fecha.timetuple()))
# sql = u"UPDATE mooc_course SET cacherev = %s WHERE id = %s" % (fecha, 133)
# cursor.execute(sql)
#
# materiasasignadas=MateriaAsignada.objects.filter(id__in=[3181412,
# 3184195,
# 3183126,
# 3186003,
# 3186514,
# 3182791,
# 3191431,
# 3192569,
# 3188785,
# 3186457,
# 3193144,
# 3196377,
# 3195969,
# 3197067,
# 3183476,
# 3186503])
# print(materiasasignadas.count())
# for eMateriaAsignada in materiasasignadas:
#     erubros = eMateriaAsignada.rubro.all()
#     if not Pago.objects.filter(rubro__in=erubros).exists():
#         erubros.delete()
#         print("ELIMINA RUBRO")
#     print("elimina maa %s"%eMateriaAsignada)
#     eMateriaAsignada.delete()


def calcular_edad(fecha_nacimiento):
    from datetime import date
    fecha_actual = date.today()
    edad = fecha_actual.year - fecha_nacimiento.year
    if fecha_actual.month < fecha_nacimiento.month or (fecha_actual.month == fecha_nacimiento.month and fecha_actual.day < fecha_nacimiento.day):
        edad -= 1
    return edad

# fecha_nacimiento = date(1990, 12, 7)
# edad = calcular_edad(fecha_nacimiento)
# print("La edad es:", edad)

def modificar_titulos(eNivel_id,hoja):
    with transaction.atomic():
        try:
            miarchivo = openpyxl.load_workbook("REGLAMENTO NOMENCLATURA DE TITULOS CES.xlsx")
            lista = miarchivo.get_sheet_by_name('%s'%hoja)
            totallista = lista.rows
            a=0
            for filas in totallista:
                a += 1
                if a > 2:
                    cod_ca = filas[0].value
                    campo_amplio = (filas[1].value).upper().replace("\n", " ")
                    cod_ce = str(filas[2].value)
                    campo_especifico = (filas[3].value).upper().replace("\n", " ")
                    cod_cd = str(filas[4].value)
                    campo_detallado = (filas[5].value).upper().replace("\n", " ")
                    cod_programa = (filas[6].value).upper().replace("\n", " ")
                    nombre_programa = (filas[7].value).upper().replace("\n", " ")
                    titulo = (filas[9].value).upper().replace("\n", " ")

                    try:
                        ePrograma=ProgramaPostulate.objects.get(nombre=nombre_programa)
                    except ObjectDoesNotExist:
                        ePrograma = ProgramaPostulate(nombre=nombre_programa,codigo=cod_programa)
                        ePrograma.save()

                    eCampoAmplio=AreaConocimientoTitulacion.objects.get(nombre=campo_amplio,tipo=1)
                    try:
                        eCampoEspecifico=SubAreaConocimientoTitulacion.objects.get(nombre=campo_especifico,tipo=1,areaconocimiento=eCampoAmplio)
                    except ObjectDoesNotExist:
                        eCampoEspecifico = SubAreaConocimientoTitulacion(nombre=campo_especifico,tipo=1,
                                                                         codigo=cod_ce,areaconocimiento=eCampoAmplio)
                        eCampoEspecifico.save()
                    try:
                        eCampoDetallado=SubAreaEspecificaConocimientoTitulacion.objects.get(nombre=campo_detallado,tipo=1,areaconocimiento=eCampoEspecifico)
                    except ObjectDoesNotExist:
                        eCampoDetallado = SubAreaEspecificaConocimientoTitulacion(nombre=campo_detallado, tipo=1,
                                                                                  codigo=cod_cd,areaconocimiento=eCampoEspecifico)
                        eCampoDetallado.save()
                    try:
                        eTitulo=Titulo.objects.get(nombre=titulo,nivel_id=eNivel_id)
                        eTitulo.usoseleccion=True
                        eTitulo.areaconocimiento=eCampoAmplio
                        eTitulo.subareaconocimiento=eCampoEspecifico
                        eTitulo.subareaespecificaconocimiento=eCampoDetallado
                        eTitulo.save()
                        campotitulo = None
                        if CamposTitulosPostulacion.objects.filter(status=True, titulo=eTitulo).exists():
                            campotitulo = CamposTitulosPostulacion.objects.filter(status=True, titulo=eTitulo).first()
                        else:
                            campotitulo = CamposTitulosPostulacion(titulo=eTitulo)
                            campotitulo.save()
                        if eTitulo.areaconocimiento:
                            if not campotitulo.campoamplio.filter(id=eTitulo.areaconocimiento.id):
                                campotitulo.campoamplio.add(eTitulo.areaconocimiento)
                        if eTitulo.subareaconocimiento:
                            if not campotitulo.campoespecifico.filter(id=eTitulo.subareaconocimiento.id):
                                campotitulo.campoespecifico.add(eTitulo.subareaconocimiento)
                        if eTitulo.subareaespecificaconocimiento:
                            if not campotitulo.campodetallado.filter(id=eTitulo.subareaespecificaconocimiento.id):
                                campotitulo.campodetallado.add(eTitulo.subareaespecificaconocimiento)
                        campotitulo.save()
                    except ObjectDoesNotExist:
                        eTitulo=Titulo(nombre=titulo,areaconocimiento=eCampoAmplio,grado_id=1,
                                       subareaconocimiento=eCampoEspecifico,nivel_id=eNivel_id,
                                       subareaespecificaconocimiento=eCampoDetallado,usoseleccion=True)
                        eTitulo.save()
                        campotitulo = None
                        if CamposTitulosPostulacion.objects.filter(status=True, titulo=eTitulo).exists():
                            campotitulo = CamposTitulosPostulacion.objects.filter(status=True, titulo=eTitulo).first()
                        else:
                            campotitulo = CamposTitulosPostulacion(titulo=eTitulo)
                            campotitulo.save()
                        if eTitulo.areaconocimiento:
                            if not campotitulo.campoamplio.filter(id=eTitulo.areaconocimiento.id):
                                campotitulo.campoamplio.add(eTitulo.areaconocimiento)
                        if eTitulo.subareaconocimiento:
                            if not campotitulo.campoespecifico.filter(id=eTitulo.subareaconocimiento.id):
                                campotitulo.campoespecifico.add(eTitulo.subareaconocimiento)
                        if eTitulo.subareaespecificaconocimiento:
                            if not campotitulo.campodetallado.filter(id=eTitulo.subareaespecificaconocimiento.id):
                                campotitulo.campodetallado.add(eTitulo.subareaespecificaconocimiento)
                        campotitulo.save()

                    try:
                        eArmonizacion=ArmonizacionNomenclaturaTitulo.objects.get(campoamplio=eCampoAmplio,
                                                                             campoespecifico=eCampoEspecifico,
                                                                             campodetallado=eCampoDetallado,
                                                                             programa=ePrograma,
                                                                             titulo=eTitulo)
                    except ObjectDoesNotExist:
                        eArmonizacion = ArmonizacionNomenclaturaTitulo(campoamplio=eCampoAmplio,
                                                                             campoespecifico=eCampoEspecifico,
                                                                             campodetallado=eCampoDetallado,
                                                                             programa=ePrograma,
                                                                             titulo=eTitulo)
                        eArmonizacion.save()
                    filas[10].value="SI"
                    print('%s - Registro titulo: %s' % (a,titulo))
            miarchivo.save("REGLAMENTO NOMENCLATURA DE TITULOS CES.xlsx")
        except Exception as ex:
            transaction.set_rollback(True)
            print('Error  {} en linea {} en registro numero {}'.format(ex, sys.exc_info()[-1].tb_lineno, str(a)))
            print('error: %s' % ex)

def ingresar_partidas():
    from postulate.models import Convocatoria, Partida
    from sga.models import Carrera, AreaConocimientoTitulacion, SubAreaConocimientoTitulacion, \
        SubAreaEspecificaConocimientoTitulacion
    a = 0
    with transaction.atomic():
        try:
            miarchivo = openpyxl.load_workbook("importar_partidas.xlsx")
            lista = miarchivo.get_sheet_by_name('ADM')
            totallista = lista.rows
            for filas in totallista:
                a += 1
                if a > 2:
                    try:
                        codigo=filas[0].value.strip()
                        print(codigo)
                        convocatoria=int(filas[18].value)
                        carr = Carrera.objects.get(nombre=filas[1].value)
                        dedicacion = 1 if filas[3].value == 'TIEMPO COMPLETO' else 2
                        rmu = 1676 if filas[3].value == 'TIEMPO COMPLETO' else 838
                        nombre_programa = (filas[4].value).upper().replace("\n", " ")
                        nombre_campocamplio=(filas[5].value).upper().replace("\n", " ")
                        nombre_campoespecifico=(filas[6].value).upper().replace("\n", " ")
                        nombre_campodetallado=(filas[7].value).upper().replace("\n", " ")
                        eTitulo_tercer = None
                        eTitulo_cuarto = None
                        ePrograma = None
                        ePrograma = ProgramaPostulate.objects.get(nombre=nombre_programa)
                        eCampoAmplio = AreaConocimientoTitulacion.objects.get(status=True, nombre=nombre_campocamplio,tipo=1)
                        eCampoEspecifico = SubAreaConocimientoTitulacion.objects.get(areaconocimiento=eCampoAmplio, status=True,nombre=nombre_campoespecifico,tipo=1)
                        eCampoDetallado = SubAreaEspecificaConocimientoTitulacion.objects.get(areaconocimiento=eCampoEspecifico, status=True,nombre=nombre_campodetallado,tipo=1)
                        try:
                            ePartida = Partida.objects.get(titulo=codigo, convocatoria_id=convocatoria)
                        except ObjectDoesNotExist:
                            ePartida = Partida(titulo=codigo, convocatoria_id=convocatoria, descripcion=filas[0].value,
                                           ano=2023, codpartida=codigo, carrera=carr, nivel=4, modalidad=5,
                                           dedicacion=dedicacion,  rmu=rmu)
                            ePartida.save()
                        if eCampoAmplio:
                            if not ePartida.campoamplio.filter(id=eCampoAmplio.id):
                                ePartida.campoamplio.add(eCampoAmplio)
                                if eCampoEspecifico:
                                    if not ePartida.campoespecifico.filter(id=eCampoEspecifico.id):
                                        ePartida.campoespecifico.add(eCampoEspecifico)
                                        if not ePartida.campodetallado.filter(id=eCampoDetallado.id):
                                            if eCampoDetallado:
                                                ePartida.campodetallado.add(eCampoDetallado)
                        eAsignatura1 = None
                        eAsignatura2 = None
                        eAsignatura3 = None
                        eAsignatura4 = None
                        eAsignatura5 = None
                        eAsignatura6 = None

                        if filas[10].value:
                            nombre_asignatura1 = (filas[10].value).upper().replace("\n", " ")
                            try:
                                eAsignatura1=Asignatura.objects.get(nombre=nombre_asignatura1)
                                if not PartidaAsignaturas.objects.filter(partida=ePartida, asignatura=eAsignatura1,
                                                                         status=True).exists():
                                    partidasave1 = PartidaAsignaturas(partida=ePartida,
                                                                     asignatura=eAsignatura1)
                                    partidasave1.save()
                            except:
                                pass
                        if filas[11].value:
                            nombre_asignatura2 = (filas[11].value).upper().replace("\n", " ")
                            try:
                                eAsignatura2 = Asignatura.objects.get(nombre=nombre_asignatura2)
                                if not PartidaAsignaturas.objects.filter(partida=ePartida, asignatura=eAsignatura2,
                                                                         status=True).exists():
                                    partidasave2 = PartidaAsignaturas(partida=ePartida,
                                                                     asignatura=eAsignatura2)
                                    partidasave2.save()
                            except:
                                pass
                        if filas[12].value:
                            nombre_asignatura3 = (filas[12].value).upper().replace("\n", " ")
                            try:
                                eAsignatura3 = Asignatura.objects.get(nombre=nombre_asignatura3)
                                if not PartidaAsignaturas.objects.filter(partida=ePartida, asignatura=eAsignatura3,
                                                                         status=True).exists():
                                    partidasave3 = PartidaAsignaturas(partida=ePartida,
                                                                     asignatura=eAsignatura3)
                                    partidasave3.save()
                            except:
                                pass
                        if filas[13].value:
                            nombre_asignatura4 = (filas[13].value).upper().replace("\n", " ")
                            try:
                                eAsignatura4 = Asignatura.objects.get(nombre=nombre_asignatura4)
                                if not PartidaAsignaturas.objects.filter(partida=ePartida, asignatura=eAsignatura4,
                                                                         status=True).exists():
                                    partidasave4 = PartidaAsignaturas(partida=ePartida,
                                                                     asignatura=eAsignatura4)
                                    partidasave4.save()
                            except:
                                pass
                        if filas[14].value:
                            nombre_asignatura5 = (filas[14].value).upper().replace("\n", " ")
                            try:
                                eAsignatura5 = Asignatura.objects.get(nombre=nombre_asignatura5)
                                if not PartidaAsignaturas.objects.filter(partida=ePartida, asignatura=eAsignatura5,
                                                                         status=True).exists():
                                    partidasave5 = PartidaAsignaturas(partida=ePartida,
                                                                     asignatura=eAsignatura5)
                                    partidasave5.save()
                            except:
                                pass
                        if filas[15].value:
                            nombre_asignatura6 = (filas[15].value).upper().replace("\n", " ")
                            try:
                                eAsignatura6 = Asignatura.objects.get(nombre=nombre_asignatura6)
                                if not PartidaAsignaturas.objects.filter(partida=ePartida, asignatura=eAsignatura6,
                                                                         status=True).exists():
                                    partidasave6 = PartidaAsignaturas(partida=ePartida,
                                                                     asignatura=eAsignatura6)
                                    partidasave6.save()
                            except:
                                pass
                        if filas[8].value:
                            titulo_tercer = (filas[8].value).upper().replace("\n", " ") if filas[8].value else None
                            eTitulo_tercer = Titulo.objects.get(nombre=titulo_tercer)
                            try:
                                eArmonizacion_tercer = ArmonizacionNomenclaturaTitulo.objects.get(campoamplio=eCampoAmplio,
                                                                                                  campoespecifico=eCampoEspecifico,
                                                                                                  campodetallado=eCampoDetallado,
                                                                                                  programa=ePrograma,
                                                                                                  titulo=eTitulo_tercer)
                                try:
                                    ePartidaArmonizacionNomenclaturaTitulo = PartidaArmonizacionNomenclaturaTitulo.objects.get(
                                        partida=ePartida, combinacion=eArmonizacion_tercer)
                                except ObjectDoesNotExist:
                                    ePartidaArmonizacionNomenclaturaTitulo = PartidaArmonizacionNomenclaturaTitulo(
                                        partida=ePartida, combinacion=eArmonizacion_tercer)
                                    ePartidaArmonizacionNomenclaturaTitulo.save()
                            except:
                                pass
                        if filas[9].value:
                            titulo_cuarto = (filas[9].value).upper().replace("\n", " ") if filas[9].value else None
                            eTitulo_cuarto = Titulo.objects.get(nombre=titulo_cuarto)
                            try:
                                eArmonizacion_cuarto = ArmonizacionNomenclaturaTitulo.objects.get(campoamplio=eCampoAmplio,
                                                                                              campoespecifico=eCampoEspecifico,
                                                                                              campodetallado=eCampoDetallado,
                                                                                              programa=ePrograma,
                                                                                              titulo=eTitulo_cuarto)
                                try:
                                    ePartidaArmonizacionNomenclaturaTitulo=PartidaArmonizacionNomenclaturaTitulo.objects.get(partida=ePartida,combinacion=eArmonizacion_cuarto)
                                except ObjectDoesNotExist:
                                    ePartidaArmonizacionNomenclaturaTitulo = PartidaArmonizacionNomenclaturaTitulo(partida=ePartida,combinacion=eArmonizacion_cuarto)
                                    ePartidaArmonizacionNomenclaturaTitulo.save()
                            except:
                                pass
                        else:
                            for armo in ArmonizacionNomenclaturaTitulo.objects.filter(status=True,programa=ePrograma,titulo__nivel_id=4):
                                try:
                                    ePartidaArmonizacionNomenclaturaTitulo = PartidaArmonizacionNomenclaturaTitulo.objects.get(
                                        partida=ePartida, combinacion=armo)
                                except ObjectDoesNotExist:
                                    ePartidaArmonizacionNomenclaturaTitulo = PartidaArmonizacionNomenclaturaTitulo(
                                        partida=ePartida, combinacion=armo)
                                    ePartidaArmonizacionNomenclaturaTitulo.save()
                    except Exception as exept:
                        filas[19].value = "%s"%exept
                        print('%s -  %s' % (a, exept))
            miarchivo.save("importar_partidas.xlsx")
        except Exception as e:
            transaction.set_rollback(True)
            print(e)
            print('Error  {} en linea {} en registro numero {}'.format(e, sys.exc_info()[-1].tb_lineno, str(a)))


def PONER_FIRMAR_MATERIAS_INGLES(id_nivel):
    try:
        from sga.models import  Nivel, Persona,ConfiguracionDocumentoEvaluaciones
        nv = Nivel.objects.filter(id__in=id_nivel)
        firma1 = Persona.objects.get(id=57920)
        firma2 = Persona.objects.get(id=29110)
        for nivel in nv:
            for mt in nivel.materia_set.filter(status=True, asignaturamalla__malla_id=353,inglesepunemi=True):
                if not ConfiguracionDocumentoEvaluaciones.objects.values().filter(materia=mt, estado=1, status=True).exists():
                    cde = ConfiguracionDocumentoEvaluaciones(materia=mt, estado=1, observacion="Ninguna", status=True)
                    cde.save()
                    if not DocumentosFirmadosEvaluaciones.objects.filter(persona_id=firma1.pk, configuraciondoc_id=cde.id, status=True, subido=False):
                        dfe1 = DocumentosFirmadosEvaluaciones(configuraciondoc_id=cde.id, status=True, subido=False, persona=firma1, cargo=u"Gerente General BUCKINGHAM ENGLISH CENTER S.A. BUCKCENTER. CONTRATO N° SIE-UNEMI-588-2023")
                        dfe1.save()
                    if not DocumentosFirmadosEvaluaciones.objects.filter(persona_id=firma2.pk, configuraciondoc_id=cde.id, status=True, subido=False):
                        dfe2 = DocumentosFirmadosEvaluaciones(configuraciondoc_id=cde.id, status=True, subido=False, persona=firma2, cargo=u"Secretaria General UNEMI")
                        dfe2.save()
                    print(u"FIRMA REGISTRADA: %s"%mt)
    except Exception as ex:
        print('error: %s' % ex)

cedulas_old = ['0915536791',
# '0925853814',
'0926402769',
'0916301633',
'1206280289',
'0921323291',
'1205897950',
'0920180783',
'0302435102',
'0925566184',
'1203704190',
'0927738146',
'0921188074',
'0925004970',
'0940121338',
'0921368890',
'0928787555',
'0925854499',
'0941150476',
'0917626350',
'0927311969',
'0924775786',
'0921598777',
'1713028916',
'0604551580',
'0917717688',
'0916368756',
'0923488316',
'0917038218',
'0922870936',
'0916785538',
'0920027471',
'0923489801',
'0922334354',
'0921366852',
'0920344793',
'0929747301',
'0918306788',
'0705377299',
'1203255623',
'0929762433',
'0921764296',
'0925567489',
'0925568941',
'0924186737',
'1314881325',
'0923791081',
'1756528384',
'0908737349',
'0914538244',
'0917712911',
'0940323496',
'0941149601',
'0915787303',
]

cedulas = ['0925004970',
'1713028916',
'0917717688',
'0922870936',
'0916785538',
'0920027471',
'0941150476',
'0921366852',
'0920344793',
'0931414361',
'0942099805',
'0923791081',
'0923488316',
'0924505258',
'0940121338',
'1203113855',
'1203255623',
'0908737349',
'0925568941',
'0924186737',
'0941344731',
'0942121708',
'0923480941',
'0917712911',
'0926844119',
'0940323496',
'0941149601',
'0928896687',
'0922334354',
'0916301633',
'1202788764',
'1756528384',
'0919619460',
'0921610408',
'0921188074',
'0804024693',
'0924775786',
'0917626350',
'0941511222',
'0940367386',
'0918566167',
'0926613076',
'0925851115',
'0918865114',
'0929857597',
'1204833923',
'0917035800',
'0916365786',
'0929749497',
'1205897950',
'0921148748',
'0911786689',
'1204180648',
'0921764296',
'0928787555',
'1206280289',
'0920180783',
'0925566184',
'0921071270',
'0942129735',
'0927738146',
'0925854499',
'0929747301',
'0921323291',
'1313034611',
'0916368756',
'1314881325',
'0914538244',
'0927311969',
'0302435102']
def PONER_MARCADAS():
    for persona in Persona.objects.filter(cedula__in=cedulas).exclude(cedula__in=cedulas_old):
        print(persona)
        fecha=convertir_fecha('22-07-2023')
        hora1=convertir_fecha_hora('22-07-2023 07:00:00')
        hora2=convertir_fecha_hora('22-07-2023 19:00:00')
        try:
            eLogDia = LogDia.objects.get(persona=persona, fecha=fecha)
        except ObjectDoesNotExist:
            eLogDia = LogDia(persona=persona,
                             fecha=fecha)
            eLogDia.save()
            try:
                eLogMarcada_1 = LogMarcada.objects.get(logdia=eLogDia, secuencia=1,time=hora1)
            except ObjectDoesNotExist:
                eLogMarcada_1 = LogMarcada(logdia=eLogDia,
                                       time=hora1,ipmarcada='.',
                                       secuencia=1)
                eLogMarcada_1.save()

            try:
                eLogMarcada_2 = LogMarcada.objects.get(logdia=eLogDia, secuencia=2,time=hora2)
            except ObjectDoesNotExist:
                eLogMarcada_2 = LogMarcada(logdia=eLogDia,
                                       time=hora2,ipmarcada='.',
                                       secuencia=1)
                eLogMarcada_2.save()
# PONER_MARCADAS()


def eliminar_matriculas_imgles():
    eMatriculas=Matricula.objects.filter(status=True, fecha_creacion__gte=datetime.now().date(), usuario_creacion_id__in=[20533,26972],
                                         inscripcion__persona__cedula__in=['0951523836',
                                                                    '0929130540',
                                                                    '2450097569',
                                                                    '0926302365',
                                                                    '0954065025',
                                                                    '0929565463',
                                                                    '0942338138',
                                                                    '0955902358',
                                                                    '0955880620',
                                                                    '0954578944',
                                                                    '0957420680',
                                                                    '0926272352',
                                                                    '0959091729',
                                                                    '0924643687',
                                                                    '0958785842',
                                                                    '0943905315',
                                                                    '0916483852',
                                                                    '0943967562',
                                                                    '0805450285',
                                                                    '0706062452',
                                                                    '0923173892',
                                                                    '1208583268',
                                                                    '0302866462',
                                                                    '0921496493',
                                                                    '0942081928',
                                                                    '0954767208',
                                                                    '1727505818',
                                                                    '0928819713',
                                                                    '0942309451',
                                                                    '0955232574',
                                                                    '1205797226',
                                                                    '0957736499',
                                                                    '0943309237',
                                                                    '0924697089',
                                                                    '2200457386',
                                                                    '0921832044',
                                                                    '0951566371',
                                                                    '1315842664',
                                                                    '1309135216',
                                                                    '0941101164',
                                                                    '0958642241',
                                                                    '1724590342',
                                                                    '0914238845',
                                                                    '0917194714',
                                                                    '0943415851',
                                                                    '0954613360',
                                                                    '0958835100',
                                                                    '2300131642',
                                                                    '0604661587',
                                                                    '0940408784',
                                                                    '0941768962',
                                                                    '0958913634',
                                                                    '0921200705',
                                                                    '0955088786',
                                                                    '1207448885',
                                                                    '0950356774',
                                                                    '0953371689',
                                                                    '0954021119',
                                                                    '1207131739',
                                                                    '0955582499',
                                                                    '0941561268',
                                                                    '0953078201',
                                                                    '0925193385',
                                                                    '0954397667',
                                                                    '0959440892',
                                                                    '2450174590',
                                                                    '0940275563',
                                                                    '0952865087',
                                                                    '0943692723',
                                                                    '0951105261',
                                                                    '1205167529',
                                                                    '0944363837',
                                                                    '1724349574',
                                                                    '0929031615',
                                                                    '1728571231',
                                                                    '0705999290',
                                                                    '0705831196',
                                                                    '1004715593',
                                                                    '0704640796',
                                                                    '0302437934',
                                                                    '0958664294',
                                                                    '0953334075',
                                                                    '0925561235',
                                                                    '2300373673',
                                                                    '0923700801',
                                                                    '0605669811',
                                                                    '0302866470',
                                                                    '1400636435',
                                                                    '0928547207',
                                                                    '0940812191',
                                                                    '1307837821',
                                                                    '0953132347',
                                                                    '1250357355',
                                                                    '0921374047',
                                                                    '0944071323',
                                                                    '0911363166',
                                                                    '0922957873',
                                                                    '0924019045',
                                                                    '0953748936',
                                                                    '0941385809',
                                                                    '0958445132',
                                                                    '0925012163',
                                                                    '0751072794',
                                                                    '0941065112',
                                                                    '1205794850',
                                                                    '0928482454',
                                                                    '0915379408',
                                                                    '0952883676',
                                                                    '1550140410',
                                                                    '0604007427',
                                                                    '1207009802',
                                                                    '1729840080',
                                                                    '1207929793',
                                                                    '0951396514',
                                                                    '0955053301',
                                                                    '0953639176',
                                                                    '0959480468',
                                                                    '0929116994',
                                                                    '0953506318',
                                                                    '0957898992',
                                                                    '0921176426',
                                                                    '1309697447',
                                                                    '0704529569',
                                                                    '0926453085',
                                                                    '0941344392'])
    print(eMatriculas.count())
    for matricula in eMatriculas:
        print(matricula)
        eMateriasAsignadas=MateriaAsignada.objects.filter(status=True, materia__asignaturamalla__malla_id=353, matricula=matricula)
        for materiaasignada in eMateriasAsignadas:
            rubros = materiaasignada.rubro.filter(status=True, observacion__icontains='INGLÉS SEPTIEMBRE 2023 - ENERO 2024')
            for rubro in rubros:
                if not rubro.pagos():
                    print(u" RUBRO ELIMINADO -- %s" % (rubro))
                    # rubro.delete()
        # matricula.delete()
# eliminar_matriculas_imgles()
def funciones_random():
    import datetime

    # Obtén la fecha actual
    fecha_actual = datetime.datetime.now()

    # Calcula la fecha del último domingo
    dias_hasta_domingo = fecha_actual.weekday()
    domingo_pasado = fecha_actual - datetime.timedelta(days=dias_hasta_domingo)

    # Crea una lista con las fechas de la última semana
    fechas_semana = []
    for i in range(7):
        fecha = domingo_pasado + datetime.timedelta(days=i)
        # Formatea la fecha en el formato deseado
        fecha_formateada = f"{fecha.day:02d} {fecha.strftime('%B')} {fecha.year}"
        # Obtiene el día de la semana
        dia_semana = fecha.strftime('%A')
        # Combina la fecha y el día de la semana
        fecha_final = f"{dia_semana.lower()} {fecha_formateada}"
        fechas_semana.append(fecha_final)

    # Imprime la lista de fechas de la última semana
    print(fechas_semana)

    import random

    dashStyle = (
        (1, u'ShortDashDot'),
        (2, u'ShortDot'),
        (3, u'Dash'),
        (4, u'ShortDash'),
        (5, u'')
    )

    # Obtén un índice aleatorio
    indice_aleatorio = random.randint(0, len(dashStyle) - 1)

    # Accede a la tupla correspondiente al índice aleatorio
    tupla_aleatoria = dashStyle[indice_aleatorio]

    # Imprime el índice y la tupla aleatoria
    print("Índice aleatorio:", indice_aleatorio)
    print("Tupla aleatoria:", tupla_aleatoria)

def xyz():
    SeguimientoTutor.objects.filter(id__in=[53042,53043,53045]).delete()


def inscribir_curso_ofimatica():
    with transaction.atomic():
        try:
            miarchivo = openpyxl.load_workbook("IMPORTAR G8.xlsx")
            lista = miarchivo.get_sheet_by_name('Hoja1')
            totallista = lista.rows
            a=0
            curso = Curso.objects.get(pk=5)
            for filas in totallista:
                a += 1
                if a >= 2:
                    cedula = filas[0].value
                    carrera = int(filas[3].value)
                    insccurso = None
                    if Inscripcion.objects.filter(persona__cedula=cedula, carrera_id=carrera, status=True).exists():
                        inscripcion = Inscripcion.objects.filter(persona__cedula=cedula, carrera_id=carrera,
                                                                 status=True).first()
                        correo = inscripcion.persona.emailinst
                        try:
                            insccurso = InscripcionCurso.objects.get(
                                Q(inscripcion=inscripcion) or Q(persona=inscripcion.persona), curso=curso,
                                status=True)
                        except ObjectDoesNotExist:
                            insccurso = InscripcionCurso(inscripcion=inscripcion, curso=curso, correo=correo)
                            insccurso.save()
                    if insccurso:
                        for x in AsignaturaCurso.objects.filter(curso=curso, status=True):
                            try:
                                asiginsccurso = AsignaturaInscripcionCurso.objects.get(asignaturacurso=x,
                                                                                       inscripcioncurso=insccurso,
                                                                                       status=True)
                            except ObjectDoesNotExist:
                                asiginsccurso = AsignaturaInscripcionCurso(asignaturacurso=x,
                                                                           inscripcioncurso=insccurso)
                                asiginsccurso.save()

                            for y in DetalleModeloEvaluativoOMA.objects.filter(modelo=curso.modeloevaluativo,
                                                                            status=True):
                                try:
                                    evaluagenerica = EvaluacionGenericaOMA.objects.get(detallemodeloevaluativo=y,
                                                                                    asignaturainscripcion=asiginsccurso,
                                                                                    status=True)
                                except ObjectDoesNotExist:
                                    evaluagenerica = EvaluacionGenericaOMA(detallemodeloevaluativo=y,
                                                                        asignaturainscripcion=asiginsccurso)
                                    evaluagenerica.save()

                    filas[4].value="SI"
                    print('%s - Registro titulo: %s' % (a,insccurso))
            miarchivo.save("IMPORTAR G8.xlsx")
        except Exception as ex:
            transaction.set_rollback(True)
            print('Error  {} en linea {} en registro numero {}'.format(ex, sys.exc_info()[-1].tb_lineno, str(a)))
            print('error: %s' % ex)

# inscribir_curso_ofimatica()

def enrolar_inscritos_curso_ofimatica():
    curso = Curso.objects.get(pk=5)
    from moodle import moodle
    tipourl = 1
    curso.crear_actualizar_estudiantes_curso_ofimatica(moodle, tipourl)

def cierre_materiaasignada():
    ins=InscripcionCurso.objects.get(id=624)
    for asigInscr in AsignaturaInscripcionCurso.objects.filter(status=True, inscripcioncurso=ins):
        asigInscr.cierre_materia_asignada()
# cierre_materiaasignada()


def cerrar_curso_ofimatica():
    data = {}
    data['curso'] = curso = Curso.objects.get(pk=6)
    print(curso)

    for alumno in curso.inscripcioncurso_set.filter(status=True,id=624):
        print(alumno.inscripcion.persona)
        print(curso.notas_de_moodle(alumno.inscripcion.persona))
        for notasmooc in curso.notas_de_moodle(alumno.inscripcion.persona):
            for asigInscr in AsignaturaInscripcionCurso.objects.filter(status=True, inscripcioncurso=alumno):
                campo = asigInscr.campo(notasmooc[1].upper().strip())
                print(campo)
                if type(notasmooc[0]) is Decimal:
                    if null_to_decimal(campo.valor) != float(notasmooc[0]):
                        print(notasmooc[0])
                        actualizar_nota_ofimatica(asigInscr.id, notasmooc[1].upper().strip(), notasmooc[0])
                        auditorianotas = AuditoriaNotasOma(evaluaciongenerica=campo, manual=False,
                                                           calificacion=notasmooc[0])
                        auditorianotas.save()
                else:
                    if null_to_decimal(campo.valor) != float(0):
                        actualizar_nota_ofimatica(asigInscr.id, notasmooc[1].upper().strip(), notasmooc[0])
                        auditorianotas = AuditoriaNotasOma(evaluaciongenerica=campo, manual=False,
                                                           calificacion=0)
                        auditorianotas.save()

    for asignaturainscripcion in AsignaturaInscripcionCurso.objects.filter(status=True, inscripcioncurso__curso=curso,inscripcioncurso_id=624):
        asignaturainscripcion.actualiza_estado()
        print(asignaturainscripcion.get_estado_display())
        asignaturainscripcion.cierre_materia_asignada()
        print(asignaturainscripcion)

    # aprobados= curso.inscripcioncurso_set.filter(status=True)
    # resultados_errores = generarCertificadoOfimatica( aprobados, data, IS_DEBUG=False)
    # print(resultados_errores)

# id_nivel=[1550]
# PONER_FIRMAR_MATERIAS_INGLES(id_nivel)


def quitar_nota_Derecho():
    for matricula in Matricula.objects.filter(id__in=[775865, 775676, 772134, 775666, 793493, 772694, 772329, 773288, 775836, 774188, 782401,
                                                      775486, 774397, 773420, 772664, 793443, 772184, 775736, 775413, 775837, 772238, 772491,
773810, 772443, 775989, 773046, 774563, 773282, 772658, 775081, 772510, 772338, 772341, 775951, 772224, 774931, 775111, 791962, 772410,
772557,
772122,
772787,
774227,
772561,
772854,
793548,
782403,
776020,
772334,
772424,
772153,
773880,
772228,
772612,
773425,
772331,
772534,
775474,
774087,
772222,
772292,
782411,
775771,
772279,
791944,
772680,
775511,
774887,
773027,
793472,
775318,
774322,
775641,
772389,
772872,
772480,
772958,
772305,
793435,
775708,
774095,
772172,
773445,
774994,
774390,
772333,
776247,
773183,
772138,
775523,
775044,
775996,
782348,
773575,
772887,
772697,
772461,
775925,
793466,
772309,
772401,
793528,
772206,
772203,
793516,
772282,
772785,
774437,
773008,
774938,
773686,
782537,
772415,
772128,
773660,
774835,
772382,
772342,
791951,
772237,
776245,
774707,
772864,
793508,
772306,
772191,
773874,
775477,
775856,
772351,
772348,
782522,
776253,
773204,
775266,
775494,
772290,
772543,
776239,
773011,
774266,
782363,
772446,
775807,
774976,
772555,
774900,
793540,
772715,
775176,
774376,
772165,
772632,
775019,
774443,
793442,
774897,
772406,
793432,
773089,
775327,
772519,
775885,
793490,
773085,
772709,
775556,
775823,
773388,
773977,
773233,
773318,
772583,
773336,
772530,
774612,
774497,
775794,
775948,
773276,
772137,
772316,
773444,
772532,
772835,
774655,
772949,
775765,
775431,
775732,
775872,
773886,
772246,
774553,
772326,
793427,
772124,
772565,
775761,
775946,
776013,
775834,
772248,
773371,
774956,
772438,
773759,
775577,
775776,
775763,
772113,
773293,
772725,
773935,
775514,
772772,
774648,
772192,
775830,
774036,
774782,
782354,
774149,
775256,
774229,
772621,
772141,
793505,
793484,
772999,
772549,
773366,
775966,
772943,
774336,
775918,
773733,
772188,
773900,
774038,
772119,
772303,
773624,
793525,
782356,
772514,
773544,
774413,
772281,
773154,
793511,
772769,
773170,
793536,
772240,
772241,
772247,
775471,
775969,
772525,
774061,
774283,
772262,
772396,
775078,
793429,
793533,
773246,
775339,
775933,
775341,
772314,
791921,
774960,
793082,
793422,
793494,
793457,
772127,
790391,
775652,
775701,
773460,
776234,
775233,
773484,
772736,
775894,
772474,
773868,
793524,
773572,
773565,
774191,
773296,
782398,
772408,
774291,
775687,
773167,
793555,
774732,
772198,
773567,
773403,
772185,
772551,
775820,
772616,
791927,
775775,
773601,
782402,
772109,
774210,
773379,
775397,
793547,
772935,
774261,
775745,
775787,
774047,
782383,
772235,
772463,
775615,
774073,
773962,
773779,
775050,
772547,
793491,
772757,
774218,
775635,
773674,
773690,
773612,
774885,
775962,
773711,
772964,
774143,
772875,
773475,
774791,
772778,
772199,
773111,
775821,
772142,
773549,
782376,
772180,
773937,
774547,
773635,
775108,
774970,
772452,
775624,
773151,
772273,
772146,
774643,
775971,
775878,
775502,
774560,
774963,
791954,
775829,
775876,
772869,
773638,
774724,
772893,
772595,
773438,
772244,
773349,
782454,
782527,
772421,
774362,
773796,
775896,
772371,
774718,
773291,
775980,
772339,
775457,
774681,
782416,
775725,
773773,
773633,
793489,
772149,
772267,
774301,
774319,
772145,
772161,
773562,
782534,
775000,
772980,
773038,
775880,
791929,
773422,
774173,
775747,
772927,
774383,
774652,
772249,
774056,
775580,
782374,
772545,
775517,
775450,
775693,
772171,
774808,
772538,
773431,
772733,
773072,
793469,
772929,
773207,
774633,
774506,
776257,
782368,
775296,
774346,
775429,
774627,
793551,
793561,
772202,
773618,
774775,
772271,
773777,
793461,
772136,
774523,
772170,
772293,
773504,
773951,
773749,
774909,
791914,
772624,
775290,
773529,
775955,
775316,
775253,
774370,
772256,
773448,
772258,
772226,
775064,
792336,
774529,
773114,
773986,
772641,
774738,
791930,
773145,
776233,
792432,
774339,
793541,
774492,
774509,
774621,
772201,
775499,
775796,
772225,
774922,
775862,
775059,
772352,
775279,
772638,
774388,
773030,
791910,
772213,
774499,
774703,
776018,
774152,
775379,
772115,
793458,
772322,
774669,
773708,
773680,
774470,
772405,
773832,
772298,
773851,
774769,
775998,
773857,
791926,
774934,
773351,
793455,
774570,
791916,
775025,
775307,
772398,
775136,
773584,
773117,
775463,
775982,
775217,
772302,
774867,
774763,
774829,
782456,
773883,
791911,
792314,
791959,
774215,
774578,
772330]):
        nota_original = 0
        nota_menor=0
        matricula.aprobado=False
        nota_menor1=0
        matricula.save()
        for materiaasignada in MateriaAsignada.objects.filter(matricula=matricula,materia__asignatura_id=4881):
            nota_original=materiaasignada.notafinal
            campo = materiaasignada.campo('EX')
            nota_menor1 = nota_original - campo.valor
            nota_menor=68-nota_menor1
            actualizar_nota_planificacion(materiaasignada.id, 'EX', nota_menor)
            auditoria=AuditoriaNotas.objects.filter(evaluaciongenerica=campo, manual=False).update(status=False)
            print(materiaasignada)
# quitar_nota_Derecho()



def cambiar_malla_modalidad_linea():
    for matricula in Matricula.objects.filter(status=True, nivel__periodo_id=317, inscripcion__carrera__modalidad=3):
        inscripcion=matricula.inscripcion
        eMallaNueva = Malla.objects.filter(carrera=inscripcion.carrera, vigente=True, validamatricula=True,
                                           modalidad=matricula.inscripcion.carrera.modalidad).order_by('-inicio').first()
        eInscripcionMallas = inscripcion.inscripcionmalla_set.filter(inscripcion=inscripcion)
        print("-----------------%s-------------------"%eInscripcionMallas.count())
        if eInscripcionMallas.values("id").exists():
            if eInscripcionMallas.count() == 1:
                eInscripcionMalla = eInscripcionMallas[0]
                if eInscripcionMalla.malla != eMallaNueva:
                    eInscripcionMalla.status = True
                    eInscripcionMalla.malla = eMallaNueva
                    eInscripcionMalla.save()
                    print("%s"%matricula)
            else:
                print("No se cambia malla: %s"%eInscripcionMallas)
# cambiar_malla_modalidad_linea()


def generar_rubros_perdida_gratuidad_1s2024_primernivel():
    from matricula.funciones import get_tipo_matricula
    matriculastodas=Matricula.objects.filter(status=True, nivel__periodo_id=317,nivelmalla_id=1)
    total=len(matriculastodas.values('id'))
    print("%s"%total)
    contador=0
    for matricula in matriculastodas:
        if matricula.inscripcion.persona.tiene_otro_titulo():
            contador+=1
            matricula.agregacion_aux(None)
            matricula.inscripcion.actualiza_estado_matricula()
            valid, msg, aData = get_tipo_matricula(None, matricula)
            if not valid:
                raise NameError(msg)
            cantidad_nivel = aData['cantidad_nivel']
            porcentaje_perdidad_parcial_gratuidad = aData['porcentaje_perdidad_parcial_gratuidad']
            cantidad_seleccionadas = aData['cantidad_seleccionadas']
            porcentaje_seleccionadas = int(round(
                Decimal((float(cantidad_nivel) * float(porcentaje_perdidad_parcial_gratuidad)) / 100).quantize(
                    Decimal('.00')), 0))
            if (cantidad_seleccionadas < porcentaje_seleccionadas):
                matricula.grupo_socio_economico(2)
            else:
                matricula.grupo_socio_economico(1)
            matricula.calcula_nivel()
            matricula.aranceldiferido = 2
            matricula.actacompromiso = None
            matricula.save()
            Rubro.objects.filter(matricula=matricula).update(status=False)
            print(matricula)
    print(contador)
# generar_rubros_perdida_gratuidad_1s2024_primernivel()

def notificar_jornada_primer_nivel(ePeriodo):
    print(f"Inicia proceso de notificación {ePeriodo.__str__()}")
    contador = 0
    matriculas=Matricula.objects.filter(status=True, nivel__periodo=ePeriodo, nivelmalla_id=1).exclude(inscripcion__carrera__modalidad=3)
    total=matriculas.count()
    for matricula in matriculas:
        contador += 1
        eInscripcion=matricula.inscripcion
        ePersona=matricula.inscripcion.persona
        if ePersona.email:
            with transaction.atomic():
                try:
                    lista=[]
                    lista.append(ePersona.email)
                    if contador<=5:
                        lista.append("jplacesc@unemi.edu.ec")
                    send_html_mail("[IMPORTANTE] UNEMI - Jornada asignada para matriculación.",
                                   "emails/email_notificacion_jornada_primernivel.html",
                                   {
                                       'sistema': u'Sistema de Gestión Académica',
                                       'fecha': datetime.now().date(),
                                       'hora': datetime.now().time(),
                                       'persona': ePersona,
                                       't': miinstitucion(),
                                       'jornada':matricula.inscripcion.sesion.nombre
                                    },
                                   lista,
                                   [],
                                   [],
                                   cuenta=CUENTAS_CORREOS[0][1])
                    print(f"{contador}/{total} -> Correo enviado ({ePersona.email}) +++++ Persona: {ePersona.__str__()} ")
                    eInscripcion.envioemail = True
                    eInscripcion.save()
                    time.sleep(2)
                except Exception as ex:
                    transaction.set_rollback(True)
                    print(f"Ocurrio un error en el envio del correo de la persona: {ePersona.__str__()}")
        else:
            print(f"{contador}/{total} -> No tiene correo la persona {ePersona.__str__()}")

# notificar_jornada_primer_nivel(Periodo.objects.get(id=317))

def generarCertificadoOfimatica():
    import pyqrcode
    data = {}
    dominio_sistema = 'http://127.0.0.1:8000'
    # dominio_sistema = 'https://sga.unemi.edu.ec'
    data["DOMINIO_DEL_SISTEMA"] = dominio_sistema
    lista_correctos = []
    lista_errores = []
    for inscripcion in InscripcionCurso.objects.filter(status=True, curso_id=12):
        try:
            data['aprobado'] = inscripcion
            temp = lambda x: remover_caracteres_tildes_unicode(remover_caracteres_especiales_unicode(x.__str__()))
            data['curso'] = curso = inscripcion.curso
            data['fecha_curso'] =fecha_curso =  fecha_certificado_curso(curso.fecha_inicio, curso.fecha_fin)
            data['hoy'] = curso.fecha_fin  + timedelta(days=1)

            qrname = 'qr_certificado_oma_' + str(inscripcion.id)
            folder = os.path.join(os.path.join(SITE_STORAGE, 'media', 'qrcode', 'omaCertificados', 'qr'))
            directory = os.path.join(os.path.join(SITE_STORAGE, 'media', 'qrcode', 'omaCertificados'))
            os.makedirs(f'{directory}/qr/', exist_ok=True)
            try:
                os.stat(directory)
            except:
                os.mkdir(directory)
            nombrepersona = temp(inscripcion.persona.__str__()).replace(' ', '_')
            htmlname = 'CERTIFICADO_OFIMATICA_{}_{}'.format(nombrepersona, random.randint(1, 100000).__str__())
            urlname = "/media/qrcode/omaCertificados/%s" % htmlname
            # rutahtml = SITE_STORAGE + urlname
            data['url_qr'] = url_qr = f'{SITE_STORAGE}/media/qrcode/omaCertificados/qr/{htmlname}.png'
            # if os.path.isfile(rutahtml):
            #     os.remove(rutahtml)
            url = pyqrcode.create(f'{dominio_sistema}/media/qrcode/omaCertificados/{htmlname}.pdf')
            imageqr = url.png(f'{directory}/qr/{htmlname}.png', 16, '#000000')
            data['qrname'] = 'qr' + qrname
            valida = conviert_html_to_pdfsaveqr_omacertificado(
                '../../oma_curso/certificado/formato_certificado.html',
                {'pagesize': 'A4', 'data': data},
                htmlname + '.pdf'
            )
            if valida:
                inscripcion.archivocertificado = 'qrcode/omaCertificados/' + htmlname + '.pdf'
                inscripcion.save()

                tituloemail = "Generación de Certificado"

                send_html_mail(tituloemail,
                               "emails/confirmacion_generacion_certificado_oma.html",
                               {'sistema': u'EPUNEMI',
                                'saludo': 'Estimada' if inscripcion.persona.sexo.id == 1 else 'Estimado',
                                'aprobado': inscripcion,
                                'curso': inscripcion.curso
                                },
                               inscripcion.lista_emails_envio_oma(),
                               [],
                               cuenta=CUENTAS_CORREOS[16][1]
                               )

                lista_correctos.append(f'{inscripcion.persona.cedula} [{inscripcion.id}]\n')
        except Exception as ex:
            lista_errores.append(
            f'{inscripcion.persona.cedula} [{inscripcion.id}] error {str(ex)} on line {str(sys.exc_info()[-1].tb_lineno)}\n')
    return lista_errores

def fecha_certificado_curso(fecha1,fecha2):
    mes = ["enero", "febrero", "marzo", "abril", "mayo", "junio", "julio", "agosto", "septiembre",
                "octubre", "noviembre", "diciembre"]

    if fecha1.month == fecha2.month:
        fecha = '{} al {} de {} de {}'.format(fecha1.day, fecha2.day, mes[fecha2.month - 1], fecha2.year)
    elif fecha1.year == fecha2.year:
        fecha = '{} de {} al {} de {} de {}'.format(fecha1.day, mes[fecha1.month-1], fecha2.day, mes[fecha2.month-1], fecha2.year)
    else:
        fecha = '{} de {} de {} al {} de {} de {}'.format(fecha1.day, mes[fecha1.month-1], fecha1.year, fecha2.day, mes[fecha2.month-1], fecha2.year)
    return fecha


# generarCertificadoOfimatica()

# variableg=VariablesGlobales.objects.get(id=105)
# variableg.valor='False'
# variableg.save()

# modulo=Modulo.objects.get(id=3)
# modulo.activo=False
# modulo.save()

# print(ProfesorMateria.objects.filter(status=True,materia__nivel__periodo_id=317, materia__status=True).distinct('profesor__persona_id').count())

def cambiar_clave_profesores():
    for profesormateria in ProfesorMateria.objects.filter(status=True,materia__nivel__periodo_id=317, materia__status=True).distinct('profesor__persona_id'):
        profesormateria.profesor.persona.cambiar_clave()
        print(profesormateria)

# cambiar_clave_profesores()

def arreglas_mis_auditorias():
    try:
        miarchivo = openpyxl.load_workbook("moodlepos_updates2.xlsx")
        lista = miarchivo.get_sheet_by_name('Hoja1')
        totallista = lista.rows
        a=0
        for filas in totallista:
            a += 1
            if a > 2:
                id = int(filas[0].value)
                query = str(filas[5].value)
                filas= str(filas[6].value)
                LogQuery.objects.filter(id=id).update(query=query,filasafectadas=filas)
                print(id)
    except Exception as ex:
            print('error: %s' % ex)
# arreglas_mis_auditorias()

def migrarMoodle():
    from Moodle_Funciones import CrearRecursoMoodle,CrearExamenMoodle,CrearCompendioMoodle,CrearMaterialesMoodle,CrearTareasMoodle,CrearTestMoodle,CrearForosMoodle
    from sga.models import Persona
    # CrearRecursoMoodle(465233,None)
    # CrearMaterialesMoodle(709278,None)
    # erecurso = CrearCompendioMoodle(133084, Persona.objects.get(id=147830))
    # print(erecurso)
    # print("Compendio creado en Moodle")

    # erecurso = CrearRecursoMoodle(452022, Persona.objects.get(id=147830))
    # print(erecurso)
    # print("Recurso creado en Moodle")
    #
    # fcurso = CrearForosMoodle(24243, Persona.objects.get(id=147830))
    # print(fcurso)
    # print("Foro creado en Moodle")
    #
    # fcurso = CrearTareasMoodle(169877, Persona.objects.get(id=147830))
    # print(fcurso)
    # print("Tarea creado en Moodle")
    #
    from django.db import connections
    ftcurso = CrearExamenMoodle(134235, Persona.objects.get(id=843))
    print(ftcurso)
    print("Test creado en Moodle")
    conexionz = connections['aulagradoa']
    cursorz = conexionz.cursor()
    sql1 = """
           select * from mooc_quiz where course=354
       """
    cursorz.execute(sql1)
    buscar = cursorz.fetchall()
    print(buscar)

migrarMoodle()
print(u"FIN")