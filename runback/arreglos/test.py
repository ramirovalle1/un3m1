import datetime
import io
import os
import sys

import pyqrcode
import xlsxwriter
import xlwt
import openpyxl
from xlwt import *

# SITE_ROOT = os.path.dirname(os.path.realpath(__file__))


YOUR_PATH = os.path.dirname(os.path.realpath(__file__))
# print(f"YOUR_PATH: {YOUR_PATH}")
SITE_ROOT = os.path.dirname(os.path.dirname(YOUR_PATH))
SITE_ROOT = os.path.join(SITE_ROOT, '')
# print(f"SITE_ROOT: {SITE_ROOT}")
your_djangoproject_home = os.path.split(SITE_ROOT)[0]
# print(f"your_djangoproject_home: {your_djangoproject_home}")
sys.path.append(your_djangoproject_home)

from django.core.wsgi import get_wsgi_application

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "settings")
application = get_wsgi_application()

from django.http import HttpResponse
from settings import MEDIA_ROOT, BASE_DIR, EMAIL_INSTITUCIONAL_AUTOMATICO, EMAIL_DOMAIN
from xlwt import easyxf, XFStyle
from sga.adm_criteriosactividadesdocente import asistencia_tutoria
from inno.models import *
from sga.models import *
from sagest.models import *
from Moodle_Funciones import buscarQuiz, estadoQuizIndividual, accesoQuizIndividual
from api.helpers.functions_helper import generate_qr_examen_final, conviert_html_to_pdf_save_path
from sga.templatetags.sga_extras import encrypt
import threading
from moodle import moodle
from sga.funciones import calculate_username, generar_usuario_admision, fechatope, llenar_requisitostitulacion, log
from idioma.models import GrupoInscripcion

profe = []


def descrypt(scr):
    try:
        print(encrypt(scr))
    except:
        print('error')


def migrar_criterios_para_admision():
    with transaction.atomic():
        try:
            distributivoadm = None
            criterios = CriterioDocenciaPeriodo.objects.filter(criterio_id__in=[141, 158, 127], periodo_id=153)

            for criterio in criterios:
                for distri in criterio.detalledistributivo_set.filter(status=True):
                    if criterio.criterio.admision:
                        profe.append(distri.distributivo.profesor.pk)
                        print(distri.distributivo.profesor)
                        periodoadm = Periodo.objects.get(id=202)
                        if not CriterioDocenciaPeriodo.objects.filter(status=True, criterio=criterio.criterio,
                                                                      periodo=periodoadm).exists():
                            criterioadm = CriterioDocenciaPeriodo(criterio=criterio.criterio, periodo=periodoadm)
                            criterioadm.save()
                        if CriterioDocenciaPeriodo.objects.filter(status=True, criterio=criterio.criterio,
                                                                  periodo=periodoadm).exists():
                            criterioadm = \
                                CriterioDocenciaPeriodo.objects.filter(status=True, criterio=criterio.criterio,
                                                                       periodo=periodoadm)[0]
                            if ProfesorDistributivoHoras.objects.filter(status=True, periodo=periodoadm,
                                                                        profesor=distri.distributivo.profesor).exists():
                                distributivoadm = ProfesorDistributivoHoras.objects.filter(status=True,
                                                                                           periodo=periodoadm,
                                                                                           profesor=distri.distributivo.profesor).first()
                            else:
                                distributivoadm = ProfesorDistributivoHoras(periodo=periodoadm,
                                                                            profesor=distri.distributivo.profesor,
                                                                            dedicacion=distri.distributivo.dedicacion,
                                                                            horasdocencia=0,
                                                                            horasinvestigacion=0,
                                                                            horasgestion=0,
                                                                            horasvinculacion=0,
                                                                            coordinacion=distri.distributivo.profesor.coordinacion,
                                                                            categoria=distri.distributivo.profesor.categoria,
                                                                            nivelcategoria=distri.distributivo.profesor.nivelcategoria,
                                                                            cargo=distri.distributivo.profesor.cargo,
                                                                            nivelescalafon=distri.distributivo.profesor.nivelescalafon)
                                distributivoadm.save()
                            if not DetalleDistributivo.objects.filter(distributivo=distributivoadm,
                                                                      criteriodocenciaperiodo=criterioadm).exists():
                                detalleadm = DetalleDistributivo(distributivo=distributivoadm,
                                                                 criteriodocenciaperiodo=criterioadm,
                                                                 criterioinvestigacionperiodo=None,
                                                                 criteriogestionperiodo=None,
                                                                 criteriovinculacionperiodo=None,
                                                                 horas=distri.horas)
                                detalleadm.save()
                                detalleadm.verifica_actividades(horas=detalleadm.horas)
                    distributivoadm.save()
                    if distributivoadm:
                        distributivoadm.resumen_evaluacion_acreditacion().actualizar_resumen()
            print(profe)
        except Exception as e:
            transaction.set_rollback(True)
            print(e)


def descarga_resultado_encuesta(id, nombre_archivo):
    response = HttpResponse(content_type="application/ms-excel")
    response['Content-Disposition'] = f'attachment; filename={nombre_archivo}.xls'
    style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
    style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
    style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
    title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
    style1 = easyxf(num_format_str='D-MMM-YY')
    font_style = XFStyle()
    font_style.font.bold = True
    font_style2 = XFStyle()
    font_style2.font.bold = False
    wb = xlwt.Workbook()
    ws = wb.add_sheet('Sheetname')
    estilo = xlwt.easyxf(
        'font: height 350, name Arial, colour_index black, bold on, italic on; align: wrap on, vert centre, horiz center;')
    ws.write_merge(0, 0, 0, 9, 'UNIVERSIDAD ESTATAL DE MILAGRO', estilo)
    output_folder = os.path.join(os.path.join(SITE_STORAGE, 'archivos'))
    try:
        os.stat(output_folder)
    except:
        os.mkdir(output_folder)
    nombre = nombre_archivo + "_" + datetime.now().strftime('%Y%m%d_%H%M%S') + ".xls"
    filename = os.path.join(output_folder, nombre)
    encuesta = EncuestaGrupoEstudiantes.objects.get(id=id)
    preguntas = encuesta.preguntaencuestagrupoestudiantes_set.filter(status=True).order_by('orden')
    columns = [(u"Nº.", 2000),
               (u"ID", 2000),
               (u"RESPONDIO", 3000),
               (u"CÉDULA", 3000),
               (u"ENCUESTADO", 9000),
               ]

    if encuesta.tipoperfil == 1:  # ALUMNO
        columns.append((u'CARRERA', 9000), )
        # solo para encuesta con y sin discapacidad
        if encuesta.id == 20 or encuesta.id == 21 or encuesta.id == 22 or encuesta.id == 23 or encuesta.id in [44,
                                                                                                               45]:  # encuestas
            columns.append((u'TIENE DISCAPACIDAD', 3000), )
            columns.append((u'EXTRANJERO', 3000), )
            columns.append((u'PPL', 3000), )
        # fin

    if encuesta.tipoperfil == 2:  # DOCENTE
        columns.append((u'Tipo de relación laboral', 9000), )
        columns.append((u'Tiempo de dedicación', 9000), )
        columns.append((u'Si es docente titular a qué categoría académica pertenece', 9000), )
        # columns.append((u'MODALIDAD CONTRATACIÓN', 9000), )
        columns.append((u'A qué Facultad pertenece', 9000), )
        columns.append((u'Las carreras en las que imparte docencia  actualmente ¿De qué modalidad son? ', 9000), )
        # columns.append((u'MODALIDAD DE LA CARRERA QUE DESEARÍA TRABAJAR EN EL SEMESTRE 1S 2022', 9000), )

    if encuesta.tipoperfil == 3:  # ADMINISTRATIVO
        columns.append((u'Tipo de relación laboral', 9000), )
        columns.append((u'Denominación del puesto', 9000), )

    for x in preguntas:
        columns.append((str(x.orden) + ") " + x.descripcion, 6000), )
        if x.tipo == 1:
            if not x.esta_vacia():
                columns.append((str(x.orden) + ") " + x.observacionporno, 6000), )
    row_num = 3
    for col_num in range(len(columns)):
        ws.write(row_num, col_num, columns[col_num][0], font_style)
        ws.col(col_num).width = columns[col_num][1]
    row_num = 4
    i = 0
    datos = []
    if encuesta.tipoperfil == 1:  # ALUMNO
        datos = encuesta.inscripcionencuestagrupoestudiantes_set.filter(status=True).order_by(
            'inscripcion__persona__apellido1')
    if encuesta.tipoperfil == 2:  # DOCENTE
        datos = encuesta.inscripcionencuestagrupoestudiantes_set.filter(status=True).order_by(
            'profesor__persona__apellido1')
    if encuesta.tipoperfil == 3:  # ADMINISTRATIVO
        datos = encuesta.inscripcionencuestagrupoestudiantes_set.filter(status=True).order_by(
            'administrativo__persona__apellido1')
    cout_register = len(datos)
    register_start = 0
    limit = 0
    for dato in datos:
        try:
            row_num += limit
            i += 1
            limit = 0
            ws.write(row_num, 0, i, font_style2)

            if encuesta.tipoperfil == 1:
                ws.write(row_num, 1, dato.inscripcion_id, font_style2)
                ws.write(row_num, 2, dato.respondio, font_style2)
                ws.write(row_num, 3, dato.inscripcion.persona.documento(), font_style2)
                ws.write(row_num, 4, dato.inscripcion.persona.nombre_completo_inverso(), font_style2)
            if encuesta.tipoperfil == 2:
                ws.write(row_num, 1, dato.profesor_id, font_style2)
                ws.write(row_num, 2, dato.respondio, font_style2)
                ws.write(row_num, 3, dato.profesor.persona.documento(), font_style2)
                ws.write(row_num, 4, dato.profesor.persona.nombre_completo_inverso(), font_style2)
            if encuesta.tipoperfil == 3:
                ws.write(row_num, 1, dato.administrativo_id, font_style2)
                ws.write(row_num, 2, dato.respondio, font_style2)
                ws.write(row_num, 3, dato.administrativo.persona.documento(), font_style2)
                ws.write(row_num, 4, dato.administrativo.persona.nombre_completo_inverso(), font_style2)
            c = 5
            if encuesta.tipoperfil == 1:
                ws.write(row_num, c, dato.inscripcion.carrera.__str__(),
                         font_style2) if not dato.inscripcion.carrera == None else ' '
                c += 1

            # solo para encuesta con y sin discapacidad
            if encuesta.id == 20 or encuesta.id == 21 or encuesta.id == 22 or encuesta.id == 23 or encuesta.id in [44,
                                                                                                                   45]:  # encuesta
                discapacidad = 'SI' if dato.inscripcion.persona.mi_perfil().tienediscapacidad else 'NO'
                ws.write(row_num, c, discapacidad, font_style2)
                c += 1

                extranjero = 'SI' if not dato.inscripcion.persona.pais_id == 1 else 'NO'
                ws.write(row_num, c, extranjero, font_style2)
                c += 1

                ppl = 'SI' if not dato.inscripcion.persona.ppl else 'NO'
                ws.write(row_num, c, ppl, font_style2)
                c += 1

            # fin

            if encuesta.tipoperfil == 2:
                dt = ProfesorDistributivoHoras.objects.filter(status=True, periodo=126,
                                                              profesor_id=dato.profesor.id).first()
                ws.write(row_num, c, dt.nivelcategoria.nombre if dt is not None else '', font_style2)
                c += 1
                ws.write(row_num, c, dt.dedicacion.nombre if dt is not None else '', font_style2)
                c += 1
                ws.write(row_num, c, dt.categoria.nombre if dt is not None and dt.nivelcategoria.id == 1 else '',
                         font_style2)
                c += 1
                ws.write(row_num, c, dt.coordinacion.nombre if dt is not None and dt.coordinacion is not None else '',
                         font_style2)
                c += 1
                w = 0
                for m in dato.profesor.mis_materias(126).values_list('materia__nivel__modalidad__nombre',
                                                                     flat=True).distinct(
                    'materia__nivel__modalidad__nombre'):
                    ws.write(row_num + w, c, str(m), font_style2)
                    w += 1
                if limit < w and w > 0:
                    limit = w - 1

                c += 1
            if encuesta.tipoperfil == 3:
                eDistributivoPersonas = DistributivoPersona.objects.filter(persona=dato.administrativo.persona,
                                                                           status=True, regimenlaboral_id=2,
                                                                           estadopuesto_id=1)
                eDistributivoPersona = None
                if eDistributivoPersonas.values("id").exists():
                    eDistributivoPersona = eDistributivoPersonas.first()
                ws.write(row_num, c,
                         eDistributivoPersona.regimenlaboral.descripcion if eDistributivoPersona is not None else '',
                         font_style2)
                c += 1
                ws.write(row_num, c,
                         eDistributivoPersona.denominacionpuesto.descripcion if eDistributivoPersona is not None else '',
                         font_style2)
                c += 1
            if dato.respondio:
                for x in preguntas:
                    respuesta = None
                    if x.tipo == 1:
                        respuesta = RespuestaPreguntaEncuestaGrupoEstudiantes.objects.filter(status=True, pregunta=x,
                                                                                             inscripcionencuesta=dato)
                        if respuesta.values("id").exists():
                            respuesta = respuesta.first()
                            ws.write(row_num, c, respuesta.respuesta, font_style2)
                            if not x.esta_vacia():
                                c += 1
                                ws.write(row_num, c, respuesta.respuestaporno, font_style2)

                        else:
                            respuesta = None
                            ws.write(row_num, c, '', font_style2)
                            if not x.esta_vacia():
                                c += 1
                                ws.write(row_num, c, '', font_style2)
                        c += 1
                    elif x.tipo == 2:
                        respuesta = RespuestaRangoEncuestaGrupoEstudiantes.objects.filter(status=True, pregunta=x,
                                                                                          inscripcionencuesta=dato)
                        if respuesta.values("id").exists():
                            respuesta = respuesta.first()
                            ws.write(row_num, c, respuesta.opcionrango.descripcion, font_style2)
                        else:
                            ws.write(row_num, c, '', font_style2)
                        c += 1
                    elif x.tipo in [3, 4]:
                        respuesta = RespuestaPreguntaEncuestaGrupoEstudiantes.objects.filter(status=True, pregunta=x,
                                                                                             inscripcionencuesta=dato)
                        if respuesta.values("id").exists():
                            respuesta = respuesta.first()
                            ws.write(row_num, c, respuesta.respuesta, font_style2)
                        else:
                            ws.write(row_num, c, '', font_style2)
                        c += 1
                    elif x.tipo == 5:
                        respuesta = dato.respuestacuadriculaencuestagrupoestudiantes_set.filter(status=True,
                                                                                                pregunta=x).first() if dato.respuestacuadriculaencuestagrupoestudiantes_set.filter(
                            status=True, pregunta=x).exists() else None
                        if respuesta is not None:
                            try:
                                int(respuesta.respuesta)
                                if OpcionCuadriculaEncuestaGrupoEstudiantes.objects.filter(status=True, pregunta=x,
                                                                                           id=respuesta.opcioncuadricula.id,
                                                                                           tipoopcion=2).first() == None:
                                    resp = 'Sin contestar'
                                else:
                                    resp = OpcionCuadriculaEncuestaGrupoEstudiantes.objects.filter(status=True,
                                                                                                   pregunta=x,
                                                                                                   id=respuesta.opcioncuadricula.id,
                                                                                                   tipoopcion=2).first().descripcion
                            except ValueError:
                                resp = respuesta.respuesta

                            ws.write(row_num, c, resp, font_style2)
                        else:
                            ws.write(row_num, c, '', font_style2)
                        c += 1
                    elif x.tipo == 6:
                        respuesta = dato.respuestamultipleencuestagrupoestudiantes_set.filter(status=True,
                                                                                              pregunta=x) if dato.respuestamultipleencuestagrupoestudiantes_set.values(
                            'id').filter(status=True, pregunta=x).exists() else None
                        if respuesta is not None:
                            w = 0
                            for rmult in respuesta:
                                ws.write(row_num + w, c, rmult.opcionmultiple.descripcion, font_style2)
                                # row_num += 1
                                w += 1
                            if limit < w and w > 0:
                                limit = w - 1
                        else:
                            ws.write(row_num, c, '', font_style2)
                        c += 1
            row_num += 1
            print('%s' % (row_num))

        except Exception as ex:
            print('error: %s' % (ex))
            pass

    wb.save(filename)
    print("FIN: ", filename)

    # if eNotificacion:
    #     eNotificacion.en_proceso = False
    #     eNotificacion.cuerpo = 'Reporte Listo'
    #     eNotificacion.url = "{}reportes/encuestas/{}".format(MEDIA_URL, nombre)
    #     eNotificacion.save()
    # else:
    #
    #     eNotificacion = Notificacion(cuerpo='Reporte Listo',
    #                                  titulo='Resultados de Encuesta',
    #                                  destinatario=ePersona,
    #                                  url="{}reportes/encuestas/{}".format(MEDIA_URL, nombre),
    #                                  prioridad=1,
    #                                  app_label='SGA',
    #                                  fecha_hora_visible=datetime.now() + timedelta(days=1),
    #                                  tipo=2,
    #                                  en_proceso=False)
    #     eNotificacion.save()
    # data = {}
    # send_user_notification(user=usernotify, payload={"head": "Reporte terminado",
    #                                                  "body": 'Resultado de Encuesta',
    #                                                  "action": "notificacion",
    #                                                  "timestamp": time.mktime(datetime.now().timetuple()),
    #                                                  "url": "{}reportes/encuestas/{}".format(MEDIA_URL, nombre_archivo),
    #                                                  "btn_notificaciones": traerNotificaciones(None, None, ePersona),
    #                                                  "mensaje": 'Los resultados de la encuesta han sido generados con exito'},
    #                        ttl=500)


# descarga_resultado_encuesta(44, 'resultados_pregrado')


# PARA CREAR TODAS LAS PLANIFICACIONES NE CADENA


def crear_matricula_admision_sede_examen_milagro_precencial():
    periodo_id = 202
    detallemodeloevaluativo_id = 114
    c = 1
    for matricula_id in Matricula.objects.filter(status=True, retiradomatricula=False,
                                                 nivel__periodo_id=periodo_id).exclude(
        inscripcion__modalidad_id=3).values_list('id', flat=True):
        print(f"Fila {c}")
        # documento = fila[col_documento].value
        # inscripcion_id = fila[col_inscripcion].value
        # eMatriculas = Matricula.objects.filter(Q(inscripcion__persona__cedula=documento) | Q(inscripcion__persona__pasaporte=documento), status=True, nivel__periodo_id=periodo_id)
        eMatriculas = Matricula.objects.filter(id=matricula_id, status=True, nivel__periodo_id=periodo_id)
        if eMatriculas.values("id").exists():
            eMatricula = eMatriculas.first()
            sede_id = 1
            with transaction.atomic():
                try:
                    eMatriculaSedeExamenes = MatriculaSedeExamen.objects.filter(matricula=eMatricula,
                                                                                detallemodeloevaluativo_id=detallemodeloevaluativo_id)
                    if not eMatriculaSedeExamenes.values("id").exists():
                        eMatriculaSedeExamen = MatriculaSedeExamen(matricula=eMatricula,
                                                                   detallemodeloevaluativo_id=detallemodeloevaluativo_id,
                                                                   sede_id=sede_id)
                        eMatriculaSedeExamen.save()
                        print(f"Se creo registro {eMatriculaSedeExamen.__str__()}")
                    # else:
                    #     eMatriculaSedeExamen = eMatriculaSedeExamenes.first()
                    #     # eMatriculaSedeExamen.sede_id = sede_id
                    #     # eMatriculaSedeExamen.save()
                    #     print(f"Se edito registro {eMatriculaSedeExamen.__str__()}")
                except Exception as ex:
                    transaction.set_rollback(True)
                    print(f"No se guardo registro, Error: {ex.__str__()}")
        else:
            print(f"No se encontro registro")
        c += 1
    print(f"********************************FINALIZA PROCESO CREAR MATRICULAS SEDE MILAGRO PRESENCIAL")

    crear_planificacion_admision_sede_examen_milagro()


def crear_matricula_admision_sede_examen_virtual():
    # pa = os.path.join(YOUR_PATH, 'archivos', 'data_encuesta_id_24.xlsx')
    folder = os.path.join(YOUR_PATH, 'archivos', 'resultados_id_45_2022_final.xlsx')
    workbook = openpyxl.load_workbook(folder)
    sheet = workbook.worksheets[2]
    all_rows = sheet.rows
    linea = 0
    col_documento = 0
    periodo_id = 202
    sede_id = 11
    detallemodeloevaluativo_id = 114
    for fila in all_rows:
        linea += 1
        if linea > 1:
            print(f"Fila {linea}")
            documento = fila[col_documento].value
            eMatriculas = Matricula.objects.filter(
                Q(inscripcion__persona__cedula=documento) | Q(inscripcion__persona__pasaporte=documento), status=True,
                nivel__periodo_id=periodo_id)
            if eMatriculas.values("id").exists():
                eMatricula = eMatriculas.first()
                with transaction.atomic():
                    try:
                        eMateriaAsignadaPlanificacionSedeVirtualExamen = MateriaAsignadaPlanificacionSedeVirtualExamen.objects.filter(
                            materiaasignada__matricula=eMatricula)
                        print(
                            f"Se procede a eliminar planificación {len(eMateriaAsignadaPlanificacionSedeVirtualExamen.values('id'))}")
                        eMateriaAsignadaPlanificacionSedeVirtualExamen.delete()
                        eMatriculaSedeExamenes = MatriculaSedeExamen.objects.filter(matricula=eMatricula,
                                                                                    detallemodeloevaluativo_id=detallemodeloevaluativo_id)
                        if not eMatriculaSedeExamenes.values("id").exists():
                            eMatriculaSedeExamen = MatriculaSedeExamen(matricula=eMatricula,
                                                                       detallemodeloevaluativo_id=detallemodeloevaluativo_id,
                                                                       sede_id=sede_id)
                            eMatriculaSedeExamen.save()
                            print(f"Se creo registro {eMatriculaSedeExamen.__str__()}")
                        else:
                            eMatriculaSedeExamen = eMatriculaSedeExamenes.first()
                            eMatriculaSedeExamen.sede_id = sede_id
                            eMatriculaSedeExamen.save()
                            print(f"Se edito registro {eMatriculaSedeExamen.__str__()}")
                    except Exception as ex:
                        transaction.set_rollback(True)
                        print(f"No se guardo registro, Error: {ex.__str__()}")
            else:
                print(f"No se encontro registro")
    print(f"********************************FINALIZA PROCESO CREAR MATRICULAS SEDE VIRTUAL")
    crear_matricula_admision_sede_examen_milagro_precencial()


def crear_matricula_admision_sede_examen_santo_domingo():
    # pa = os.path.join(YOUR_PATH, 'archivos', 'data_encuesta_id_24.xlsx')
    folder = os.path.join(YOUR_PATH, 'archivos', 'resultados_id_45_2022_final.xlsx')
    workbook = openpyxl.load_workbook(folder)
    sheet = workbook.worksheets[1]
    all_rows = sheet.rows
    linea = 0
    # col_documento = 0
    col_inscripcion = 0
    col_sede_id = 1
    periodo_id = 202
    detallemodeloevaluativo_id = 114
    for fila in all_rows:
        linea += 1
        if linea > 1:
            print(f"Fila {linea}")
            # documento = fila[col_documento].value
            inscripcion_id = fila[col_inscripcion].value
            # eMatriculas = Matricula.objects.filter(Q(inscripcion__persona__cedula=documento) | Q(inscripcion__persona__pasaporte=documento), status=True, nivel__periodo_id=periodo_id)
            eMatriculas = Matricula.objects.filter(inscripcion_id=inscripcion_id, status=True,
                                                   nivel__periodo_id=periodo_id)
            if eMatriculas.values("id").exists():
                eMatricula = eMatriculas.first()
                sede_id = int(fila[col_sede_id].value)
                with transaction.atomic():
                    try:
                        eMatriculaSedeExamenes = MatriculaSedeExamen.objects.filter(matricula=eMatricula,
                                                                                    detallemodeloevaluativo_id=detallemodeloevaluativo_id)
                        if not eMatriculaSedeExamenes.values("id").exists():
                            eMatriculaSedeExamen = MatriculaSedeExamen(matricula=eMatricula,
                                                                       detallemodeloevaluativo_id=detallemodeloevaluativo_id,
                                                                       sede_id=sede_id)
                            eMatriculaSedeExamen.save()
                            print(f"Se creo registro {eMatriculaSedeExamen.__str__()}")
                        else:
                            eMatriculaSedeExamen = eMatriculaSedeExamenes.first()
                            # eMatriculaSedeExamen.sede_id = sede_id
                            # eMatriculaSedeExamen.save()
                            print(f"Se edito registro {eMatriculaSedeExamen.__str__()}")
                    except Exception as ex:
                        transaction.set_rollback(True)
                        print(f"No se guardo registro, Error: {ex.__str__()}")
            else:
                print(f"No se encontro registro")
    print(f"********************************FINALIZA PROCESO CREAR MATRICULAS SEDE SANTO DOMINGO")
    crear_matricula_admision_sede_examen_virtual()


def crear_matricula_admision_sede_examen_milagro():
    folder = os.path.join(YOUR_PATH, 'archivos', 'resultados_id_45_2022_final.xlsx')
    # folder = os.path.join(os.path.join(BASE_DIR, 'archivos', 'resultados_id_45_2022_final.xlsx'))
    workbook = openpyxl.load_workbook(folder)
    sheet = workbook.worksheets[0]
    all_rows = sheet.rows
    linea = 0
    # col_documento = 0
    col_inscripcion = 0
    col_sede_id = 1
    periodo_id = 202
    detallemodeloevaluativo_id = 114
    for fila in all_rows:
        linea += 1
        if linea > 1:
            print(f"Fila {linea}")
            # documento = fila[col_documento].value
            inscripcion_id = fila[col_inscripcion].value
            # eMatriculas = Matricula.objects.filter(Q(inscripcion__persona__cedula=documento) | Q(inscripcion__persona__pasaporte=documento), status=True, nivel__periodo_id=periodo_id)
            eMatriculas = Matricula.objects.filter(inscripcion_id=inscripcion_id, status=True,
                                                   nivel__periodo_id=periodo_id)
            if eMatriculas.values("id").exists():
                eMatricula = eMatriculas.first()
                sede_id = int(fila[col_sede_id].value)
                with transaction.atomic():
                    try:
                        eMatriculaSedeExamenes = MatriculaSedeExamen.objects.filter(matricula=eMatricula,
                                                                                    detallemodeloevaluativo_id=detallemodeloevaluativo_id)
                        if not eMatriculaSedeExamenes.values("id").exists():
                            eMatriculaSedeExamen = MatriculaSedeExamen(matricula=eMatricula,
                                                                       detallemodeloevaluativo_id=detallemodeloevaluativo_id,
                                                                       sede_id=sede_id)
                            eMatriculaSedeExamen.save()
                            print(f"Se creo registro {eMatriculaSedeExamen.__str__()}")
                        else:
                            eMatriculaSedeExamen = eMatriculaSedeExamenes.first()
                            eMatriculaSedeExamen.sede_id = sede_id
                            eMatriculaSedeExamen.save()
                            print(f"Se edito registro {eMatriculaSedeExamen.__str__()}")
                    except Exception as ex:
                        transaction.set_rollback(True)
                        print(f"No se guardo registro, Error: {ex.__str__()}")
            else:
                print(f"No se encontro registro")
    print(f"********************************FINALIZA PROCESO CREAR MATRICULAS SEDE MILAGRO")
    crear_matricula_admision_sede_examen_santo_domingo()


def planificar_admision_unemi_v1(sede):
    sede_id = sede
    periodo_id = 202
    detallemodeloevaluativo_id = 114
    # if DEBUG:
    #     MateriaAsignadaPlanificacionSedeVirtualExamen.objects.filter(status=True,
    #                                                                  aulaplanificacion__turnoplanificacion__fechaplanificacion__periodo_id=periodo_id,
    #                                                                  detallemodeloevaluativo_id=detallemodeloevaluativo_id).delete()
    eMatriculaSedeExamenes = MatriculaSedeExamen.objects.filter(sede_id=sede_id,
                                                                detallemodeloevaluativo_id=detallemodeloevaluativo_id,
                                                                status=True, matricula__status=True,
                                                                matricula__retiradomatricula=False,
                                                                matricula__nivel__periodo_id=periodo_id,
                                                                matricula__inscripcion__carrera__id__in=(
                                                                    223, 224)).distinct()
    # eMallaIngles = Malla.objects.filter(pk__in=[353, 22])
    # eMateriaAsignadaPlanificacionSedeVirtualExamenes = MateriaAsignadaPlanificacionSedeVirtualExamen.objects.filter(aulaplanificacion__turnoplanificacion__fechaplanificacion__periodo_id=periodo_id, aulaplanificacion__turnoplanificacion__fechaplanificacion__sede_id=sede_id)
    # print(f"Se procedera a eliminar {len(eMateriaAsignadaPlanificacionSedeVirtualExamenes)}")
    # eMateriaAsignadaPlanificacionSedeVirtualExamenes.delete()

    # eMateriaAsignadaPlanificacionSedeVirtualExamenestot = MateriaAsignadaPlanificacionSedeVirtualExamen.objects.filter(
    #     status=True, aulaplanificacion__turnoplanificacion__fechaplanificacion__periodo_id=periodo_id,
    #     detallemodeloevaluativo_id=detallemodeloevaluativo_id)
    #
    # eMatriculastot = Matricula.objects.filter(pk__in=eMatriculaSedeExamenes.values_list("matricula__id", flat=True),
    #                                        status=True, retiradomatricula=False, termino=True,
    #                                        fechatermino__isnull=False, bloqueomatricula=False,
    #                                        nivel__periodo_id=periodo_id)
    # eMatriculas_planificadas_tot = eMatriculastot.annotate(total_planificadas=Count('materiaasignada__id', filter=Q(
    #     materiaasignada__id__in=eMateriaAsignadaPlanificacionSedeVirtualExamenestot.values_list("materiaasignada__id",
    #                                                                                          flat=True), status=True),
    #                                                                                  nivel__periodo_id=periodo_id,
    #                                                                                  status=True),
    #                                                         total_general=Count('materiaasignada__id',
    #                                                                             filter=Q(nivel__periodo_id=periodo_id,
    #                                                                                      status=True), exclude=Q(
    #                                                                 materiaasignada__materia__asignatura__id=4837))).filter(
    #     total_general=F('total_planificadas'))
    eFechaPlanificacionSedeVirtualExamenes = FechaPlanificacionSedeVirtualExamen.objects.filter(sede_id=sede_id,
                                                                                                periodo_id=periodo_id,
                                                                                                fecha__gte='2023-02-24').order_by(
        'fecha')
    for eFechaPlanificacionSedeVirtualExamen in eFechaPlanificacionSedeVirtualExamenes:
        fecha = eFechaPlanificacionSedeVirtualExamen.fecha
        eTurnoPlanificacionSedeVirtualExamenes = TurnoPlanificacionSedeVirtualExamen.objects.filter(
            fechaplanificacion=eFechaPlanificacionSedeVirtualExamen).order_by('horainicio')
        for eTurnoPlanificacionSedeVirtualExamen in eTurnoPlanificacionSedeVirtualExamenes:
            horainicio = eTurnoPlanificacionSedeVirtualExamen.horainicio
            horafin = eTurnoPlanificacionSedeVirtualExamen.horafin
            eAulaPlanificacionSedeVirtualExamenes = AulaPlanificacionSedeVirtualExamen.objects.filter(
                turnoplanificacion=eTurnoPlanificacionSedeVirtualExamen)
            for eAulaPlanificacionSedeVirtualExamen in eAulaPlanificacionSedeVirtualExamenes:
                eAula = eAulaPlanificacionSedeVirtualExamen.aula
                capacidad = eAula.capacidad
                cantidadad_planificadas = eAulaPlanificacionSedeVirtualExamen.cantidadad_planificadas()
                if cantidadad_planificadas < capacidad:
                    print(
                        f"Se procede a planificar en la fecha {fecha} en el horario {horainicio} a {horafin} en el aula {eAula.nombre}")
                    eMateriaAsignadaPlanificacionSedeVirtualExamenes = MateriaAsignadaPlanificacionSedeVirtualExamen.objects.filter(
                        status=True, aulaplanificacion__turnoplanificacion__fechaplanificacion__periodo_id=periodo_id,
                        detallemodeloevaluativo_id=detallemodeloevaluativo_id)
                    filter_conflicto = (Q(aulaplanificacion__turnoplanificacion__horainicio__lte=horafin,
                                          aulaplanificacion__turnoplanificacion__horafin__gte=horafin,
                                          aulaplanificacion__turnoplanificacion__fechaplanificacion__fecha=fecha) |
                                        Q(aulaplanificacion__turnoplanificacion__horainicio__lte=horainicio,
                                          aulaplanificacion__turnoplanificacion__horafin__gte=horainicio,
                                          aulaplanificacion__turnoplanificacion__fechaplanificacion__fecha=fecha))
                    eMatriculas = Matricula.objects.filter(
                        pk__in=eMatriculaSedeExamenes.values_list("matricula__id", flat=True),
                        status=True, retiradomatricula=False, termino=True,
                        fechatermino__isnull=False, bloqueomatricula=False,
                        nivel__periodo_id=periodo_id)
                    # eMatriculas = eMatriculas.exclude(pk__in=eMateriaAsignadaPlanificacionSedeVirtualExamenes.values_list("materiaasignada__matricula__id", flat=True).filter(filter_conflicto))
                    # eMatriculas_exclude_ingles = eMatriculas.annotate(total_ingles=Count('materiaasignada__materia__asignaturamalla__malla_id', filter=Q(materiaasignada__materia__asignaturamalla__malla__in=eMallaIngles, nivel__periodo_id=periodo_id, status=True)),
                    #                                                   total_general=Count('materiaasignada__materia__asignaturamalla__malla_id', filter=Q(nivel__periodo_id=periodo_id, status=True))).filter(total_general=F('total_ingles'))
                    eMatriculas_exclude_planificadas = eMatriculas.annotate(
                        total_planificadas=Count('materiaasignada__id', filter=Q(
                            materiaasignada__id__in=eMateriaAsignadaPlanificacionSedeVirtualExamenes.values_list(
                                "materiaasignada__id", flat=True), status=True), nivel__periodo_id=periodo_id,
                                                 status=True),
                        total_general=Count('materiaasignada__id', filter=Q(nivel__periodo_id=periodo_id, status=True),
                                            exclude=Q(materiaasignada__materia__asignatura__id=4837))).filter(
                        total_general=F('total_planificadas'))
                    # eMatriculas = eMatriculas.exclude(pk__in=eMatriculas_exclude_ingles.values_list('id', flat=True))
                    # eMatriculas = eMatriculas.exclude(pk__in=eMatriculas_exclude_planificadas.values_list('id', flat=True))
                    ids_exclude = list(
                        eMateriaAsignadaPlanificacionSedeVirtualExamenes.values_list("materiaasignada__matricula__id",
                                                                                     flat=True).filter(
                            filter_conflicto))
                    # ids_exclude.extend(list(eMatriculas_exclude_ingles.values_list('id', flat=True)))
                    ids_exclude.extend(list(eMatriculas_exclude_planificadas.values_list('id', flat=True)))
                    eMatriculas = eMatriculas.exclude(pk__in=ids_exclude)
                    eMatriculas = eMatriculas.order_by('inscripcion__persona__apellido1',
                                                       'inscripcion__persona__apellido2',
                                                       'inscripcion__persona__nombres').distinct()
                    # eMatriculas = eMatriculas.exclude(Q(materiaasignada__id__in=eMateriaAsignadaPlanificacionSedeVirtualExamenes.values_list("materiaasignada__id", flat=True)) | Q(materiaasignada__materia__asignatura__id=4837))
                    contador = cantidadad_planificadas
                    for eMatricula in eMatriculas:
                        eMateriaAsignadas = MateriaAsignada.objects.filter(status=True, matricula=eMatricula).exclude(
                            Q(pk__in=eMateriaAsignadaPlanificacionSedeVirtualExamenes.values_list("materiaasignada__id",
                                                                                                  flat=True)) | Q(
                                materia__asignatura__id=4837))
                        if eMateriaAsignadas.values("id").exists():
                            eMateriaAsignada = eMateriaAsignadas.first()
                            contador += 1
                            print(
                                f"------- ({contador}) Se asignada el estudiante {eMateriaAsignada.matricula.inscripcion.persona} en la asignatura {eMateriaAsignada.materia.asignatura.nombre}")
                            eMateriaAsignadaPlanificacionSedeVirtualExamen = MateriaAsignadaPlanificacionSedeVirtualExamen(
                                aulaplanificacion=eAulaPlanificacionSedeVirtualExamen,
                                materiaasignada=eMateriaAsignada,
                                detallemodeloevaluativo_id=detallemodeloevaluativo_id)
                            eMateriaAsignadaPlanificacionSedeVirtualExamen.save()
                            if contador >= capacidad:
                                break


# planificar_admision_unemi_v1(1)

# crear_planificacion_admision_sede_examen_santo_domingo_tsachilas()


def inscripciones_resultados(array):
    response = HttpResponse(content_type="application/ms-excel")
    response['Content-Disposition'] = f'attachment; filename=inscripciones_2.xls'
    style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
    style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
    style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
    title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
    style1 = easyxf(num_format_str='D-MMM-YY')
    font_style = XFStyle()
    font_style.font.bold = True
    font_style2 = XFStyle()
    font_style2.font.bold = False
    wb = xlwt.Workbook()
    ws = wb.add_sheet('Sheetname')
    estilo = xlwt.easyxf(
        'font: height 350, name Arial, colour_index black, bold on, italic on; align: wrap on, vert centre, horiz center;')
    ws.write_merge(0, 0, 0, 9, 'UNIVERSIDAD ESTATAL DE MILAGRO', estilo)
    output_folder = os.path.join(os.path.join(SITE_STORAGE, 'archivos'))
    try:
        os.stat(output_folder)
    except:
        os.mkdir(output_folder)
    nombre = 'inscripciones_2' + "_" + datetime.now().strftime('%Y%m%d_%H%M%S') + ".xls"
    filename = os.path.join(output_folder, nombre)
    columns = [(u"Nº.", 2000),
               (u"ID", 2000),
               (u"CÉDULA", 3000),
               (u"ENCUESTADO", 9000),
               (u"TIENE DISCAPACIDAD", 3000),
               (u"EXTRANJERO", 3000),
               (u"PPL", 3000),
               ]
    row_num = 1
    for col_num in range(len(columns)):
        ws.write(row_num, col_num, columns[col_num][0], font_style)
        ws.col(col_num).width = columns[col_num][1]
    row_num = 5
    datos = Inscripcion.objects.filter(status=True, id__in=array).order_by('inscripcion__persona__apellido1')
    i = 0
    for dato in datos:
        try:
            i += 1
            limit = 0
            ws.write(row_num, 0, i, font_style2)
            ws.write(row_num, 1, dato.id, font_style2)
            ws.write(row_num, 3, dato.persona.documento(), font_style2)
            ws.write(row_num, 4, dato.persona.nombre_completo_inverso(), font_style2)
            ws.write(row_num, 5, dato.carrera.__str__(), font_style2) if not dato.carrera == None else ' '

            discapacidad = 'SI' if dato.persona.mi_perfil().tienediscapacidad else 'NO'
            ws.write(row_num, 6, discapacidad, font_style2)
            extranjero = 'SI' if not dato.persona.pais_id == 1 else 'NO'
            ws.write(row_num, 7, extranjero, font_style2)
            ppl = 'SI' if not dato.persona.ppl else 'NO'
            ws.write(row_num, 8, ppl, font_style2)
            row_num += 1
            print('%s' % (row_num))

        except Exception as ex:
            print('error: %s' % (ex))
            pass

    wb.save(filename)
    print("FIN: ", filename)


def asignar_tutorias():
    periodo = Periodo.objects.get(id=177)
    horarios = HorarioTutoriaAcademica.objects.filter(status=True, periodo=periodo).distinct()
    fecha_i = convertir_fecha('24-04-2023')
    fecha_f = convertir_fecha('30-04-2023')
    while fecha_i <= fecha_f:
        for horario in horarios:
            fecha = fecha_i
            if horario.dia == fecha.isoweekday():
                print(horario.profesor.persona)
                if not RegistroClaseTutoriaDocente.objects.filter(horario=horario,
                                                                  numerosemana=fecha.isocalendar()[1],
                                                                  fecha__date=fecha).exists():
                    clasetutoria = RegistroClaseTutoriaDocente(horario=horario,
                                                               numerosemana=fecha.isocalendar()[1],
                                                               fecha=fecha)
                    clasetutoria.save()
                    print('Se ingresa tutoria: %s' % clasetutoria)
        fecha_i = fecha_i + timedelta(days=1)


def inscribir_alumnos_en_sede():
    try:
        planificar_admision_unemi_v1(1)
        n = 'MILAGRO'
        print('FIN DE PLANIFICACION DE LA SEDE ' + n)
    except Exception as e:
        print(e)


def habilitar_examen_admision():
    with transaction.atomic():
        try:
            eFechaPlanificacionSedeVirtualExamenes = FechaPlanificacionSedeVirtualExamen.objects.filter(periodo_id=202,
                                                                                                        fecha='2023-03-03',
                                                                                                        sede_id=11).order_by(
                'sede', 'fecha')
            for eFechaPlanificacionSedeVirtualExamen in eFechaPlanificacionSedeVirtualExamenes:
                fecha = eFechaPlanificacionSedeVirtualExamen.fecha
                # print(eFechaPlanificacionSedeVirtualExamen.sede)
                # print(fecha)
                eTurnoPlanificacionSedeVirtualExamenes = TurnoPlanificacionSedeVirtualExamen.objects.filter(
                    fechaplanificacion=eFechaPlanificacionSedeVirtualExamen).order_by('horainicio')
                for eTurnoPlanificacionSedeVirtualExamen in eTurnoPlanificacionSedeVirtualExamenes:
                    eAulaPlanificacionSedeVirtualExamenes = AulaPlanificacionSedeVirtualExamen.objects.filter(
                        turnoplanificacion=eTurnoPlanificacionSedeVirtualExamen)
                    for eAulaPlanificacionSedeVirtualExamenid in eAulaPlanificacionSedeVirtualExamenes:
                        # eAula = eAulaPlanificacionSedeVirtualExamen.aula
                        horainicio = eTurnoPlanificacionSedeVirtualExamen.horainicio
                        horafin = eTurnoPlanificacionSedeVirtualExamen.horafin
                        # fecha = form.cleaned_data['fecha']
                        # horainicio = form.cleaned_data['horainicio']
                        # horafin = form.cleaned_data['horafin']
                        # password = form.cleaned_data['password']
                        fechadesde = datetime(fecha.year, fecha.month, fecha.day, horainicio.hour, horainicio.minute,
                                              horainicio.second)
                        fechadesde = int(time.mktime(fechadesde.timetuple()))
                        fechahasta = datetime(fecha.year, fecha.month, fecha.day, horafin.hour, horafin.minute,
                                              horafin.second)
                        fechahasta = int(time.mktime(fechahasta.timetuple()))
                        eAulaPlanificacionSedeVirtualExamen = AulaPlanificacionSedeVirtualExamen.objects.get(
                            pk=eAulaPlanificacionSedeVirtualExamenid.pk)
                        # eAulaPlanificacionSedeVirtualExamen.password = password
                        eAulaPlanificacionSedeVirtualExamen.registrohabilitacion = datetime.now()
                        eAulaPlanificacionSedeVirtualExamen.save()
                        # print(horainicio)
                        # print(eAulaPlanificacionSedeVirtualExamen.aula)
                        # if eAulaPlanificacionSedeVirtualExamen.password == '' or not eAulaPlanificacionSedeVirtualExamen.password:
                        #     raise NameError(u"Generar contraseña para hablitar el examen")
                        eMateriaAsignadaPlanificacionSedeVirtualExamenes = MateriaAsignadaPlanificacionSedeVirtualExamen.objects.filter(
                            aulaplanificacion=eAulaPlanificacionSedeVirtualExamen,
                            aulaplanificacion__turnoplanificacion__fechaplanificacion__periodo_id=202, status=True)
                        # if not eMateriaAsignadaPlanificacionSedeVirtualExamenes.values("id").exists():
                        #     raise NameError(u"No existe alumno o alumnos con registro de asistencia")
                        for eMateriaAsignadaPlanificacionSedeVirtualExamen in eMateriaAsignadaPlanificacionSedeVirtualExamenes:
                            # if not eMateriaAsignadaPlanificacionSedeVirtualExamen.habilitadoexamen:
                            eMateriaAsignada = eMateriaAsignadaPlanificacionSedeVirtualExamen.materiaasignada
                            username = eMateriaAsignada.matricula.inscripcion.persona.usuario.username
                            # print(eMateriaAsignada.matricula.inscripcion.persona)
                            # print(eMateriaAsignada.matricula.inscripcion.persona.cedula)
                            password = eMateriaAsignadaPlanificacionSedeVirtualExamen.password
                            if not eMateriaAsignadaPlanificacionSedeVirtualExamen.password or password == '' or password == None:
                                eMateriaAsignadaPlanificacionSedeVirtualExamen.create_update_password()
                                password = eMateriaAsignadaPlanificacionSedeVirtualExamen.password
                            examenplanificado = eMateriaAsignada.materia.examenplanificadosilabo(
                                eMateriaAsignadaPlanificacionSedeVirtualExamen.detallemodeloevaluativo)
                            if examenplanificado:
                                quiz = buscarQuiz(examenplanificado.get("idtestmoodle"),
                                                  eMateriaAsignada.materia.coordinacion().id)
                                limite = int(quiz[3])
                                try:
                                    intentos = int(quiz[4])
                                    if intentos == 0:
                                        intentos = examenplanificado.get('vecesintento')
                                        if intentos is None or intentos == 0:
                                            intentos = 1
                                except Exception as exIn:
                                    intentos = 1
                                estado_examen = estadoQuizIndividual(username, eMateriaAsignada.materia,
                                                                     examenplanificado.get("idtestmoodle"))
                                if estado_examen != 'inprogress':
                                    eMateriaAsignadaPlanificacionSedeVirtualExamen.idtestmoodle = int(
                                        examenplanificado.get("idtestmoodle"))
                                    eMateriaAsignadaPlanificacionSedeVirtualExamen.password = password
                                    eMateriaAsignada = eMateriaAsignadaPlanificacionSedeVirtualExamen.materiaasignada
                                    isResult, msgError = accesoQuizIndividual(
                                        eMateriaAsignada.matricula.inscripcion.persona.usuario.username,
                                        eMateriaAsignada.materia,
                                        eMateriaAsignadaPlanificacionSedeVirtualExamen.idtestmoodle,
                                        {'timeopen': fechadesde,
                                         'timeclose': fechahasta,
                                         'timelimit': limite,
                                         'password': password,
                                         'attempts': intentos})
                                    if isResult:
                                        eMateriaAsignadaPlanificacionSedeVirtualExamen.habilitadoexamen = True
                                    eMateriaAsignadaPlanificacionSedeVirtualExamen.save()
                                    print('ALUMNO ASIGNADO {}'.format(
                                        eMateriaAsignadaPlanificacionSedeVirtualExamen.materiaasignada))
                            else:
                                print('***************************************************')
                                print('***************************************************')
                                print('***************************************************')
                                print('***************************************************')
                                print('MATERIA NO PLANFICADA')
                                print(eMateriaAsignada.materia)
                                print('***************************************************')
                                print('***************************************************')
                                print('***************************************************')
                                print('***************************************************')

        except Exception as ex:
            transaction.set_rollback(True)
            print(ex)


def verificar_estado_estudiante(array):
    try:
        datos = Inscripcion.objects.filter(status=True, id__in=array).order_by('inscripcion__persona__apellido1')
        mil = []
        virt = []
        for dato in datos:
            discapacidad = 'SI' if dato.persona.mi_perfil().tienediscapacidad else 'NO'
            extranjero = 'SI' if not dato.persona.pais_id == 1 else 'NO'
            ppl = 'SI' if dato.persona.ppl else 'NO'
            if discapacidad == 'SI' or extranjero == 'SI' or ppl == 'SI':
                virt.append(dato.pk)
                # print('{} - discapacidad: {} - extranjero: {} - ppl: {}'.format(dato.persona.nombre_completo_minus(), discapacidad, extranjero, ppl))
            else:
                mil.append(dato.pk)
                # print('{} - discapacidad: {} - extranjero: {} - ppl: {}'.format(dato.persona.nombre_completo_minus(),
                #                                                         discapacidad, extranjero, ppl))
                # print('{} {} - Extranjero'.format(dato.persona.nombre_completo_minus()))
        print('SEDE MILAGRO')
        for m in mil:
            print('{}, '.format(m))
        print('SEDE VIRTUAL')
        for v in virt:
            print('{}, '.format(v))

    except Exception as e:
        print(e)


def verificar_estudiante(array):
    try:
        datos = Inscripcion.objects.filter(status=True, id__in=array).order_by('inscripcion__persona__apellido1')
        mil = []
        virt = []
        for dato in datos:
            pass
        print('SEDE MILAGRO')
        for m in mil:
            print('{}, '.format(m))
        print('SEDE VIRTUAL')
        for v in virt:
            print('{}, '.format(v))

    except Exception as e:
        print(e)


def set_ppl_statatus():
    for ppl in HistorialPersonaPPL.objects.filter(status=True):
        if not ppl.persona.ppl:
            print(ppl.persona)
        ppl.save()
    print('FINALIZA ACTUALZACION DE PPLS')


# EXAMENES PREGRADO


def crear_matricula_pregrado_sede_examen_virtual():
    # pa = os.path.join(YOUR_PATH, 'archivos', 'data_encuesta_id_24.xlsx')
    folder = os.path.join(YOUR_PATH, 'archivos', 'resultados_id_45_2022_final.xlsx')
    workbook = openpyxl.load_workbook(folder)
    sheet = workbook.worksheets[2]
    all_rows = sheet.rows
    linea = 0
    col_documento = 0
    periodo_id = 153
    sede_id = 11
    detallemodeloevaluativo_id = 37
    for fila in all_rows:
        linea += 1
        if linea > 1:
            print(f"Fila {linea}")
            documento = fila[col_documento].value
            eMatriculas = Matricula.objects.filter(
                Q(inscripcion__persona__cedula=documento) | Q(inscripcion__persona__pasaporte=documento), status=True,
                nivel__periodo_id=periodo_id)
            if eMatriculas.values("id").exists():
                eMatricula = eMatriculas.first()
                with transaction.atomic():
                    try:
                        eMateriaAsignadaPlanificacionSedeVirtualExamen = MateriaAsignadaPlanificacionSedeVirtualExamen.objects.filter(
                            materiaasignada__matricula=eMatricula)
                        print(
                            f"Se procede a eliminar planificación {len(eMateriaAsignadaPlanificacionSedeVirtualExamen.values('id'))}")
                        eMateriaAsignadaPlanificacionSedeVirtualExamen.delete()
                        eMatriculaSedeExamenes = MatriculaSedeExamen.objects.filter(matricula=eMatricula,
                                                                                    detallemodeloevaluativo_id=detallemodeloevaluativo_id)
                        if not eMatriculaSedeExamenes.values("id").exists():
                            eMatriculaSedeExamen = MatriculaSedeExamen(matricula=eMatricula,
                                                                       detallemodeloevaluativo_id=detallemodeloevaluativo_id,
                                                                       sede_id=sede_id)
                            eMatriculaSedeExamen.save()
                            print(f"Se creo registro {eMatriculaSedeExamen.__str__()}")
                        else:
                            eMatriculaSedeExamen = eMatriculaSedeExamenes.first()
                            eMatriculaSedeExamen.sede_id = sede_id
                            eMatriculaSedeExamen.save()
                            print(f"Se edito registro {eMatriculaSedeExamen.__str__()}")
                    except Exception as ex:
                        transaction.set_rollback(True)
                        print(f"No se guardo registro, Error: {ex.__str__()}")
            else:
                print(f"No se encontro registro")
    print(f"********************************FINALIZA PROCESO CREAR MATRICULAS SEDE VIRTUAL")
    crear_matricula_admision_sede_examen_milagro_precencial()


def crear_matricula_pregrado_sede_examen_santo_domingo():
    # pa = os.path.join(YOUR_PATH, 'archivos', 'data_encuesta_id_24.xlsx')
    folder = os.path.join(YOUR_PATH, 'archivos', 'resultados_id_45_2022_final.xlsx')
    workbook = openpyxl.load_workbook(folder)
    sheet = workbook.worksheets[1]
    all_rows = sheet.rows
    linea = 0
    # col_documento = 0
    col_inscripcion = 0
    col_sede_id = 1
    periodo_id = 156
    detallemodeloevaluativo_id = 37
    for fila in all_rows:
        linea += 1
        if linea > 1:
            print(f"Fila {linea}")
            # documento = fila[col_documento].value
            inscripcion_id = fila[col_inscripcion].value
            # eMatriculas = Matricula.objects.filter(Q(inscripcion__persona__cedula=documento) | Q(inscripcion__persona__pasaporte=documento), status=True, nivel__periodo_id=periodo_id)
            eMatriculas = Matricula.objects.filter(inscripcion_id=inscripcion_id, status=True,
                                                   nivel__periodo_id=periodo_id)
            if eMatriculas.values("id").exists():
                eMatricula = eMatriculas.first()
                sede_id = int(fila[col_sede_id].value)
                with transaction.atomic():
                    try:
                        eMatriculaSedeExamenes = MatriculaSedeExamen.objects.filter(matricula=eMatricula,
                                                                                    detallemodeloevaluativo_id=detallemodeloevaluativo_id)
                        if not eMatriculaSedeExamenes.values("id").exists():
                            eMatriculaSedeExamen = MatriculaSedeExamen(matricula=eMatricula,
                                                                       detallemodeloevaluativo_id=detallemodeloevaluativo_id,
                                                                       sede_id=sede_id)
                            eMatriculaSedeExamen.save()
                            print(f"Se creo registro {eMatriculaSedeExamen.__str__()}")
                        else:
                            eMatriculaSedeExamen = eMatriculaSedeExamenes.first()
                            # eMatriculaSedeExamen.sede_id = sede_id
                            # eMatriculaSedeExamen.save()
                            print(f"Se edito registro {eMatriculaSedeExamen.__str__()}")
                    except Exception as ex:
                        transaction.set_rollback(True)
                        print(f"No se guardo registro, Error: {ex.__str__()}")
            else:
                print(f"No se encontro registro")
    print(f"********************************FINALIZA PROCESO CREAR MATRICULAS SEDE SANTO DOMINGO")
    crear_matricula_admision_sede_examen_virtual()


virtual = ['1105072290',
           '0928947464',
           '0706065687',
           '0705905412',
           '0950144592',
           '0927950402',
           '0202084174',
           '0911388197',
           '1721190906',
           '1206061598',
           '0951272665',
           '0940446172',
           '0941600322',
           '1204628075',
           '0924012743',
           '0929745776',
           '0924308356',
           '2300825722',
           '1207229236',
           '0925851503',
           '0942079898',
           '1500760853',
           '0927873745',
           '0302862743',
           '0951338300',
           '1750940981',
           '0942460452',
           '0803270925',
           '0301769428',
           '0928194885',
           '0944243724',
           '0928263458',
           '0923703441',
           '1725619710',
           '0932262595',
           '1204926354',
           '0940172539',
           '0929830289',
           '0706123502',
           '0929746311',
           '0952141588',
           '0909689937',
           '0928262971',
           '0931258834',
           '0941270803',
           '1311537953',
           '0953868635',
           '0928519834',
           '0606274512',
           '0941600728',
           '0940351158',
           '0941200404',
           '0922999750',
           '0926280629',
           '1721249470',
           '0706689205',
           '0931587166',
           '0930637822',
           '0913246765',
           '0929817302',
           '0916866585',
           '0941118465',
           '0909078594',
           '0920814381',
           '0750873796',
           '0705842185',
           '1205898032',
           '0105891816',
           '0929601243',
           '0929604460',
           '0950108654',
           '0927955450',
           '0922337365',
           '0922337365',
           '0503048878',
           '0705604866',
           '0920784261',
           '0917930208',
           '0957643778',
           '0941527293',
           '0941527293',
           '1206713701',
           '0957152911',
           '1721765749',
           '0921840914',
           '1726099433',
           '1313454462',
           '0706238029',
           '0927150920',
           '0951002054',
           '2400007809',
           '1724451743',
           '1751602143',
           '1311328049',
           '0923907877',
           '0927998260',
           '0955199716',
           '0504504358',
           '0106231830',
           '0929512531',
           '1313200568',
           '0504437435',
           '0250129780',
           '0922829635',
           '2450655952',
           '0927836072',
           '1720989514',
           '1716587124',
           '0920895927',
           '0920895927',
           '1003441498',
           '0925121972',
           '0956186134',
           '1207207166',
           '1206513465',
           '0953370418',
           '1205632027',
           '0705749463',
           '0927312744',
           '1714461462',
           '0953253192',
           '1207789189',
           '0927282590',
           '0930580360',
           '0942795139',
           '0942254673',
           '0941495624',
           '1314333608',
           '0706486313',
           '0803129535',
           '0944013572',
           '1716198187',
           '0929791002',
           '0941600140',
           '2400326977',
           '0925833360',
           '1722032875',
           '1715122022',
           '0930874748',
           '0915192678',
           '0940367618',
           '0929368264',
           '1205623356',
           '0921119079',
           '0910128594',
           '1724099260',
           '1204026767',
           '1711480861',
           '1723975775',
           '0923011241',
           '1104636210',
           '1723893770',
           '1003858402',
           '0919070821',
           '0914483086',
           '1004089981',
           '0958672313',
           '0932186810',
           '1805240833',
           '0913252441',
           '0958948499',
           '0915742050',
           '0928267715',
           '0202177838',
           '1207955111',
           '1207117498',
           '0920710134',
           '1311307001',
           '0940742299',
           '1723854277',
           '0916399850',
           '0920771581',
           '0954583480',
           '0705420883',
           '1207000181',
           '0803441955',
           '1313706887',
           '0940669781',
           '0917924250',
           '0915242366',
           '1727069005',
           '0931393227',
           '2450840562',
           '0918496779',
           '0951728872',
           '0940169006',
           '1718641424',
           '1207137058',
           '0951028232',
           '0929214286',
           '0604013714',
           '0925719718',
           '0928778349',
           '0921674891',
           '0919412726',
           '0926613076',
           '0955331343',
           '0706330024',
           '1104852056',
           '0926408089',
           '0921254611',
           '1726006248',
           '0958307704',
           '0920191103',
           '0929159515',
           '0503643942',
           '0921878310',
           '1251310437',
           '0917184574',
           '0926681578',
           '0707068961',
           '0951048644',
           '0914214416',
           '0922592282',
           '1207005958',
           '0803422898',
           '1104615040',
           '1206669911',
           '1717688301',
           '0929857530',
           '0942054891',
           '0803557917',
           '0912913704',
           '0921371332',
           '0943034652',
           '0941016313',
           '0926579921',
           '0941933509',
           '0951663269',
           '1306900604',
           '1729752442',
           '1750337022',
           '1723927693',
           '1206707273',
           '0706333952',
           '1105177479',
           '0929817963',
           '0928049386',
           '0928067792',
           '0926402611',
           '1207096361',
           '0929837946',
           '0927738617',
           '0953072584',
           '0941158461',
           '1311286650',
           '1725949885',
           '1724997711',
           '1709611279',
           '0803663285',
           '1206621748',
           '2300475205',
           '0951354562',
           '0202379285',
           '0926921826',
           '0930176839',
           '0941443996',
           '1207543610',
           '0917275489',
           '0951276112',
           '2300191232',
           '0957934110',
           '0926720996',
           '0950327718',
           '1722760145',
           '0917824583',
           '0941529042',
           '0706770690',
           '0941584013',
           '0940418312',
           '0941252462',
           '0706185162',
           '0602036089',
           '0917845620',
           '0940192339',
           '1204932501',
           '1205766866',
           '0916588478',
           '1726845637',
           '0919332502',
           '1207472885',
           '1208718385',
           '0922791694',
           '0941226524',
           '0941226862',
           '0930450309',
           '1204936403',
           '0702191495',
           '0928050137',
           '0918710674',
           '0602534265',
           '0929315638',
           '0950265884',
           '0302431119',
           '1717970956',
           '0919205559',
           '0953464302',
           '0917626632',
           '0907910376',
           '0502708910',
           '0921581955',
           '0301248837',
           '0929322022',
           '0928253087',
           '0914082433',
           '0941326167',
           '0941333445',
           '1719137836',
           '1315149698',
           '0950952986',
           '1716672041',
           '0951132174',
           '0929722072',
           '0914916861',
           '0929420602',
           '0941099335',
           '0914235478',
           '2150196927',
           '0940529860',
           '1315645620',
           '1312267535',
           '0802051441',
           '0106152564',
           '0918624206',
           '0952297828',
           '0929608420',
           '0940149727',
           '1750857748',
           '1722986005',
           '0950665034',
           '0503097883',
           '0920160736',
           '0941366288',
           '0941525966',
           '0202411161',
           '0943271999',
           '0928369040',
           '1718477662',
           '0942262619',
           '0601869258',
           '0955391446',
           '1310200009',
           '0957290430',
           '1714846605',
           '1900485424',
           '0706392230',
           '1722458500',
           '0924185978',
           '0954394854',
           '0957597289',
           '0940239684',
           '0201521895',
           '0924245517',
           '0940383557',
           '0940383565',
           '1105572869',
           '1206256917',
           '1206840397',
           '0942486861',
           '1202124705',
           '0705685816',
           '0941603789',
           '1719057794',
           '1204852014',
           '0921661625',
           '0928532399',
           '0951733658',
           '0950420406',
           '0919435487',
           '1206062166',
           '0702716481',
           '0928050996',
           '0925005910',
           '0927041418',
           '0802195131',
           '0941216632',
           '0917287104',
           '0951783133',
           '0931556021',
           '0705945350',
           '0106293566',
           '0955035852',
           '0952341600',
           '0706706520',
           '1310478456',
           '1004163323',
           '1104798259',
           '1308981800',
           '0952499846',
           '1104641533',
           '0106877855',
           '0926626490',
           '0922965637',
           '2300061955',
           '2450085226',
           '1718964024',
           '0958341604',
           '1726795204',
           '0703968198',
           '0941142564',
           '0953458601',
           '1725276347',
           '1105597585',
           '0954875449',
           '0928360056',
           '0921148045',
           '0916637432',
           '0704352863',
           '0750653149',
           '0952880094',
           '0940128705',
           '0605933381',
           '0202132189',
           '0944210947',
           '0918478603',
           '1207573302',
           '0930549936',
           '0922672365',
           '0920501988',
           '0706792348',
           '1204774762',
           '0940028806',
           '1250026414',
           '0926318361',
           '0950970475',
           '0942074055',
           '0950306134',
           '0942533670',
           '0925061509',
           '1309535860',
           '0201523669',
           '0951482231',
           '0803272392',
           '0916093636',
           '0940113301',
           '0941532442',
           '0603450206',
           '0921997763',
           '0703645838',
           '0920860426',
           '0928408855',
           '0804499671',
           '0952237220',
           '0915707111',
           '2450468034',
           '1719410936',
           '1206698282',
           '0916458680',
           '2350023186',
           '0930715420',
           '0802417899',
           '0919997148',
           '0924536113',
           '1723428528',
           '0919144972',
           '2300367287',
           '2450422577',
           '0940588452',
           '0955346150',
           '0704380062',
           '1721833653',
           '0956412407',
           '1317377727',
           '1311868622',
           '0604020339',
           '1724055668',
           '0952916211',
           '0958185480',
           '0603438474',
           '0926444308',
           '1312038910',
           '0957064769',
           '1711976934',
           '0751023623',
           '0958424228',
           '0953311917',
           '0750248841',
           '0929637577',
           '0951567833',
           '0102577814',
           '0916188543',
           '0916188535',
           '0916188766',
           '0913962338',
           '0911388197',
           '0924037328',
           '0929852457',
           '0940318785',
           '0957635907',
           '1350523690',
           '1759991159',
           '0922873146',
           '0951960129',
           '0927543595',
           '1312585662',
           '1312585662',
           '0603265414',
           '0705177418',
           '0951141266',
           '0605386986',
           '0918623612',
           '0930798335',
           '0925019796',
           '0963273297',
           '1205466715',
           '0705172666',
           '1805169826',
           '2350578213',
           '1719558510',
           '0954811998',
           '0928174754',
           '1600077372',
           '0916309404',
           '1805274063',
           '0941158305',
           '0941158305',
           '1713088514',
           '0940088966',
           '1804853677',
           '0950524405',
           '1207137074',
           '0922592555',
           'AY018110',
           '2100344775',
           '1205238650',
           '0923168017',
           '1751347830',
           '1721114443',
           '0922281209',
           '1600596942',
           '0921792826',
           '1004021778',
           '1804764809',
           '0801588260',
           '1720543212',
           '0953828456',
           '1204956229',
           '0750975880',
           '0602763617',
           '1726135906',
           '0921350823',
           '0958361958',
           '0926691080',
           '1103689798',
           '1723660294',
           '1753480274',
           '1105676173',
           '1205452673',
           '0705953628',
           '0924712987',
           '0705025575',
           '0924536113',
           '0929393247',
           '0703130302',
           '0922902028',
           '1751139344',
           '0926739210',
           '1709611279',
           '0931772529',
           '1205919945',
           '1350214233',
           '0928646702',
           '0925803363',
           '0952035517',
           '1805082391',
           '1104318686',
           '2300364144',
           '0923711246',
           '1206472415',
           '0604189332',
           '0942103177',
           '0931064554',
           '0940761604',
           '0917279291',
           '1722419114',
           '1311885774',
           '1309151478',
           '1719589283',
           '0925375560',
           '1804421541',
           '0940735046',
           '1719710988',
           '1104172513',
           '1206542365',
           '0930566088',
           '0955950159',
           '2400313991',
           '0103722492',
           '0944010479',
           '0707042586',
           '2100765581',
           '0302208590',
           '0928526573',
           '0705144186',
           '0503340820',
           '0940187016',
           '0925007973',
           '2200071476',
           '0942252552',
           '1805094966',
           '0950794479',
           '0924443112',
           '1802093854',
           '1725083867',
           '0922512256',
           '1719681833',
           '1718663808',
           '1316824844',
           '0965123896',
           '0928364124',
           '0956564637',
           '0958172033',
           '0927213587',
           '0924010713',
           '0920324506']


def crear_matricula_pregrado_sede_examen_milagro():
    folder = os.path.join(YOUR_PATH, 'archivos', 'resultados_pregrado.xlsx')
    # folder = os.path.join(os.path.join(BASE_DIR, 'archivos', 'resultados_id_45_2022_final.xlsx'))
    workbook = openpyxl.load_workbook(folder)
    sheet = workbook.worksheets[0]
    all_rows = sheet.rows
    linea = 0
    # col_documento = 0
    col_inscripcion = 1
    col_sede_id = 10
    periodo_id = 153
    detallemodeloevaluativo_id = 37
    for fila in all_rows:
        linea += 1
        if linea > 1:
            print(f"Fila {linea}")
            # documento = fila[col_documento].value
            inscripcion_id = fila[col_inscripcion].value
            # eMatriculas = Matricula.objects.filter(Q(inscripcion__persona__cedula=documento) | Q(inscripcion__persona__pasaporte=documento), status=True, nivel__periodo_id=periodo_id)
            eMatriculas = Matricula.objects.filter(inscripcion_id=inscripcion_id, status=True,
                                                   nivel__periodo_id=periodo_id)
            if eMatriculas.values("id").exists():
                eMatricula = eMatriculas.first()
                discapacidad = eMatricula.inscripcion.persona.mi_perfil().tienediscapacidad = eMatricula.inscripcion.persona.mi_perfil().tienediscapacidad
                extranjero = True if not eMatricula.inscripcion.persona.pais_id == 1 else False
                ppl = eMatricula.inscripcion.persona.ppl
                sede_id = fila[col_sede_id].value
                if sede_id == 'MILAGRO':
                    sede_id = 1
                elif sede_id == 'SANTO DOMINGO':
                    sede_id = 10
                # elif sede_id == 'EXTRANJERO' or discapacidad or extranjero or ppl:
                #     sede_id = 11
                else:
                    sede_id = 1
                if discapacidad or extranjero or ppl or eMatricula.inscripcion.persona.cedula in virtual:
                    sede_id = 11
                with transaction.atomic():
                    try:
                        eMatriculaSedeExamenes = MatriculaSedeExamen.objects.filter(matricula=eMatricula,
                                                                                    detallemodeloevaluativo_id=detallemodeloevaluativo_id)
                        if not eMatriculaSedeExamenes.values("id").exists():
                            eMatriculaSedeExamen = MatriculaSedeExamen(matricula=eMatricula,
                                                                       detallemodeloevaluativo_id=detallemodeloevaluativo_id,
                                                                       sede_id=sede_id)
                            eMatriculaSedeExamen.save()
                            print(f"Se creo registro {eMatriculaSedeExamen.__str__()}")
                        else:
                            eMatriculaSedeExamen = eMatriculaSedeExamenes.first()
                            eMatriculaSedeExamen.sede_id = sede_id
                            eMatriculaSedeExamen.save()
                            print(f"Se edito registro {eMatriculaSedeExamen.__str__()}")
                    except Exception as ex:
                        transaction.set_rollback(True)
                        print(f"No se guardo registro, Error: {ex.__str__()}")
            else:
                print(f"No se encontro registro")
    print(f"********************************FINALIZA PROCESO CREAR MATRICULAS SEDE MILAGRO")
    crear_planificacion_pregrado_sede_examen_santo_domingo_tsachilas()


def mover_alumnos():
    for ppl in MatriculaSedeExamen.objects.filter(
            Q(matricula__inscripcion__persona__ppl=True, matricula__nivel__periodo_id=202) | Q(
                matricula__inscripcion__persona__cedula__in=['1803993516',
                                                             '1803966264',
                                                             '0703612275',
                                                             '0605225481',
                                                             '0924773351',
                                                             '0802000638',
                                                             '0202041828',
                                                             '0603684028',
                                                             '1850588045',
                                                             '1003165808']
            )):
        # update inno_materiaasignadaplanificacionsedevirtualexamen set aulaplanificacion_id=21776
        list1 = [27183, 27182] if DEBUG else [21998, 21999]
        materias = MateriaAsignadaPlanificacionSedeVirtualExamen.objects.filter(
            materiaasignada__matricula=ppl.matricula)
        for materia in materias:
            r = random.choice(list1)
            list1.remove(r)
            while not materias.filter(aulaplanificacion_id=r).exists():
                materia.aulaplanificacion_id = r
                materia.save()
        ppl.sede_id = 11
        ppl.save()
    habilitar_examen_admision()


def check(sede_id):
    re = []

    matriculas = MatriculaSedeExamen.objects.filter(sede_id=sede_id, detallemodeloevaluativo_id=114,
                                                    status=True, matricula__status=True,
                                                    matricula__retiradomatricula=False,
                                                    matricula__nivel__periodo_id=177).distinct()
    for matricula in matriculas:
        matricula_id = matricula.matricula.id
        materiasasi = MateriaAsignada.objects.filter(status=True, matricula=matricula_id).exclude(
            materia__asignaturamalla__malla_id__in=Malla.objects.filter(pk__in=[353, 22]).values_list('id',
                                                                                                      flat=True)).values(
            'id')
        planificadas = MateriaAsignadaPlanificacionSedeVirtualExamen.objects.filter(
            materiaasignada__matricula=matricula_id).exclude(
            materiaasignada__materia__asignaturamalla__malla_id__in=Malla.objects.filter(pk__in=[353, 22]).values_list(
                'id', flat=True)).values('id')
        if len(materiasasi) > len(planificadas):
            print(matricula_id)
    return re


def planificar_pregrado_unemi_v1(sede, limite_materia):
    try:
        sede_id = sede
        periodo_id = 177
        detallemodeloevaluativo_id = 114
        eMallasIngles = Malla.objects.filter(pk__in=[353, 22]).values_list('id', flat=True)
        c = len(MateriaAsignadaPlanificacionSedeVirtualExamen.objects.filter(status=True,
                                                                             aulaplanificacion__turnoplanificacion__fechaplanificacion__periodo_id=periodo_id,
                                                                             detallemodeloevaluativo_id=detallemodeloevaluativo_id).values(
            'id'))
        # if DEBUG:
        comp = []
        exlall = []
        # MateriaAsignadaPlanificacionSedeVirtualExamen.objects.filter(status=True,
        #                                                                  aulaplanificacion__turnoplanificacion__fechaplanificacion__periodo_id=periodo_id,
        #                                                                  detallemodeloevaluativo_id=detallemodeloevaluativo_id).delete()
        eMatriculaSedeExamenes = MatriculaSedeExamen.objects.filter(sede_id=sede_id,
                                                                    detallemodeloevaluativo_id=detallemodeloevaluativo_id,
                                                                    status=True, matricula__status=True,
                                                                    matricula__retiradomatricula=False,
                                                                    matricula__nivel__periodo_id=periodo_id).distinct()
        eFechaPlanificacionSedeVirtualExamenes = FechaPlanificacionSedeVirtualExamen.objects.filter(sede_id=sede_id,
                                                                                                    periodo_id=periodo_id).order_by(
            'fecha')
        eMatriculas = Matricula.objects.filter(pk__in=eMatriculaSedeExamenes.values_list("matricula__id", flat=True),
                                               status=True, retiradomatricula=False, bloqueomatricula=False,
                                               nivel__periodo_id=periodo_id)
        eMatriculas_exclude_ingles = eMatriculas.annotate(
            total_ingles=Count('materiaasignada__materia__asignaturamalla__id', filter=Q(
                materiaasignada__materia__asignaturamalla__malla_id__in=eMallasIngles.values_list('id', flat=True),
                nivel__periodo_id=periodo_id, status=True)),
            total_general=Count('materiaasignada__materia__asignaturamalla__id',
                                filter=Q(nivel__periodo_id=periodo_id, status=True))).filter(
            total_general=F('total_ingles'))
        ids_exclude = []
        ids_exclude.extend(list(eMatriculas_exclude_ingles.values_list('id', flat=True)))
        eMatriculas = eMatriculas.exclude(pk__in=ids_exclude)
        eMallas = Malla.objects.filter(
            pk__in=eMatriculas.values_list('inscripcion__inscripcionmalla__malla_id', flat=True).distinct()).exclude(
            id__in=eMallasIngles)
        eMallas = eMallas.order_by('carrera__nombre', 'inicio')
        for malla in eMallas:
            for eFechaPlanificacionSedeVirtualExamen in eFechaPlanificacionSedeVirtualExamenes:
                fecha = eFechaPlanificacionSedeVirtualExamen.fecha
                eTurnoPlanificacionSedeVirtualExamenes = TurnoPlanificacionSedeVirtualExamen.objects.filter(
                    fechaplanificacion=eFechaPlanificacionSedeVirtualExamen).order_by('horainicio')
                for eTurnoPlanificacionSedeVirtualExamen in eTurnoPlanificacionSedeVirtualExamenes:
                    horainicio = eTurnoPlanificacionSedeVirtualExamen.horainicio
                    horafin = eTurnoPlanificacionSedeVirtualExamen.horafin
                    eAulaPlanificacionSedeVirtualExamenes = AulaPlanificacionSedeVirtualExamen.objects.filter(
                        turnoplanificacion=eTurnoPlanificacionSedeVirtualExamen)
                    for eAulaPlanificacionSedeVirtualExamen in eAulaPlanificacionSedeVirtualExamenes:
                        eAula = eAulaPlanificacionSedeVirtualExamen.aula
                        capacidad = eAula.capacidad
                        cantidadad_planificadas = eAulaPlanificacionSedeVirtualExamen.cantidadad_planificadas()
                        if cantidadad_planificadas < capacidad:
                            print(
                                f"Se procede a planificar en la fecha {fecha} en el horario {horainicio} a {horafin} en el aula {eAula.nombre}")
                            eMateriaAsignadaPlanificacionSedeVirtualExamenes = MateriaAsignadaPlanificacionSedeVirtualExamen.objects.filter(
                                status=True,
                                aulaplanificacion__turnoplanificacion__fechaplanificacion__periodo_id=periodo_id,
                                detallemodeloevaluativo_id=detallemodeloevaluativo_id).values_list(
                                'materiaasignada__id', flat=True).distinct()
                            excludematriculas = MateriaAsignadaPlanificacionSedeVirtualExamen.objects.filter(
                                status=True,
                                aulaplanificacion__turnoplanificacion__fechaplanificacion__periodo_id=periodo_id,
                                detallemodeloevaluativo_id=detallemodeloevaluativo_id).values_list(
                                'materiaasignada__matricula__id', flat=True).distinct()
                            filter_conflicto = (Q(aulaplanificacion__turnoplanificacion__horainicio__lte=horafin,
                                                  aulaplanificacion__turnoplanificacion__horafin__gte=horafin,
                                                  aulaplanificacion__turnoplanificacion__fechaplanificacion__fecha=fecha) |
                                                Q(aulaplanificacion__turnoplanificacion__horainicio__lte=horainicio,
                                                  aulaplanificacion__turnoplanificacion__horafin__gte=horainicio,
                                                  aulaplanificacion__turnoplanificacion__fechaplanificacion__fecha=fecha))
                            eMatriculas = Matricula.objects.filter(
                                pk__in=eMatriculaSedeExamenes.values_list("matricula__id", flat=True).exclude(
                                    matricula__id__in=exlall),
                                status=True, retiradomatricula=False, termino=True,
                                fechatermino__isnull=False, bloqueomatricula=False,
                                nivel__periodo_id=periodo_id, inscripcion__inscripcionmalla__malla=malla)
                            # eMatriculas = eMatriculas.exclude(pk__in=eMateriaAsignadaPlanificacionSedeVirtualExamenes.values_list("materiaasignada__matricula__id", flat=True).filter(filter_conflicto))
                            # eMatriculas_exclude_ingles = eMatriculas.annotate(total_ingles=Count('materiaasignada__materia__asignaturamalla__malla_id', filter=Q(materiaasignada__materia__asignaturamalla__malla__in=eMallaIngles, nivel__periodo_id=periodo_id, status=True)),
                            #                                                   total_general=Count('materiaasignada__materia__asignaturamalla__malla_id', filter=Q(nivel__periodo_id=periodo_id, status=True))).filter(total_general=F('total_ingles'))
                            eMatriculas_exclude_planificadas = eMatriculas.annotate(
                                total_planificadas=Count('materiaasignada__id', filter=Q(
                                    materiaasignada__id__in=eMateriaAsignadaPlanificacionSedeVirtualExamenes,
                                    status=True), nivel__periodo_id=periodo_id,
                                                         status=True),
                                total_general=Count('materiaasignada__id',
                                                    filter=Q(nivel__periodo_id=periodo_id, status=True),
                                                    exclude=Q(materiaasignada__materia__asignatura__id=4837))).filter(
                                total_general=F('total_planificadas'))
                            # eMatriculas = eMatriculas.exclude(pk__in=eMatriculas_exclude_ingles.values_list('id', flat=True))
                            # eMatriculas = eMatriculas.exclude(pk__in=eMatriculas_exclude_planificadas.values_list('id', flat=True))

                            ids_exclude = list(excludematriculas.filter(filter_conflicto))
                            # ids_exclude.extend(list(eMatriculas_exclude_ingles.values_list('id', flat=True)))
                            ids_exclude.extend(list(eMatriculas_exclude_planificadas.values_list('id', flat=True)))
                            eMatriculas = eMatriculas.exclude(pk__in=ids_exclude)
                            ids = list(MateriaAsignadaPlanificacionSedeVirtualExamen.objects.filter(
                                aulaplanificacion__turnoplanificacion__fechaplanificacion__fecha=eFechaPlanificacionSedeVirtualExamen.fecha).values(
                                'aulaplanificacion__turnoplanificacion__fechaplanificacion__fecha').annotate(
                                total=Count('aulaplanificacion__turnoplanificacion__fechaplanificacion__fecha')).values(
                                'total').filter(total__gte=limite_materia).values_list('materiaasignada__matricula_id',
                                                                                       flat=True))

                            # if horainicio >= datetime.now().replace(2023,1,1,14,0,0,0).time() and comp:
                            #     ids.extend(comp)
                            eMatriculas = eMatriculas.exclude(pk__in=ids)
                            eMatriculas = eMatriculas.exclude(
                                pk__in=eMatriculas_exclude_ingles.values_list('id', flat=True))
                            eMatriculas = eMatriculas.order_by('inscripcion__persona__apellido1',
                                                               'inscripcion__persona__apellido2',
                                                               'inscripcion__persona__nombres').distinct()
                            # eMatriculas = eMatriculas.exclude(Q(materiaasignada__id__in=eMateriaAsignadaPlanificacionSedeVirtualExamenes.values_list("materiaasignada__id", flat=True)) | Q(materiaasignada__materia__asignatura__id=4837))
                            contador = cantidadad_planificadas
                            for fmatricula in comp:
                                if not fmatricula in exlall:
                                    if contador >= capacidad:
                                        break
                                    delet = MateriaAsignadaPlanificacionSedeVirtualExamen.objects.filter(
                                        status=True,
                                        materiaasignada__matricula_id=fmatricula).exclude(
                                        Q(materiaasignada__materia__asignatura__id=4837) |
                                        Q(materiaasignada__materia__asignaturamalla__malla_id__in=eMallasIngles.values_list(
                                            "id", flat=True))).values('id')
                                    tmaterias = MateriaAsignada.objects.filter(status=True,
                                                                               matricula_id=fmatricula).values('id')
                                    if len(delet) == len(tmaterias) and not fmatricula in exlall:
                                        comp.remove(fmatricula)
                                        exlall.append(fmatricula)
                                    if fmatricula in comp:
                                        eMateriaAsignadas = MateriaAsignada.objects.filter(status=True,
                                                                                           matricula_id=fmatricula).exclude(
                                            Q(pk__in=delet.values_list("materiaasignada__id", flat=True)) | Q(
                                                materia__asignatura__id=4837))
                                        if eMateriaAsignadas.values("id").exists():
                                            eMateriaAsignada = eMateriaAsignadas.first()
                                            if not MateriaAsignadaPlanificacionSedeVirtualExamen.objects.filter(
                                                    status=True,
                                                    materiaasignada=eMateriaAsignada,
                                                    aulaplanificacion__turnoplanificacion__horainicio=eTurnoPlanificacionSedeVirtualExamen.horainicio,
                                                    aulaplanificacion__turnoplanificacion__fechaplanificacion=eFechaPlanificacionSedeVirtualExamen,
                                                    detallemodeloevaluativo_id=detallemodeloevaluativo_id).exists():

                                                contador += 1
                                                c += 1
                                                print(
                                                    f"-------({c}) ->  ({contador}) Se asignada el estudiante {eMateriaAsignada.matricula.inscripcion.persona} en la asignatura {eMateriaAsignada.materia.asignatura.nombre}")
                                                eMateriaAsignadaPlanificacionSedeVirtualExamen = MateriaAsignadaPlanificacionSedeVirtualExamen(
                                                    aulaplanificacion=eAulaPlanificacionSedeVirtualExamen,
                                                    materiaasignada=eMateriaAsignada,
                                                    detallemodeloevaluativo_id=detallemodeloevaluativo_id)
                                                eMateriaAsignadaPlanificacionSedeVirtualExamen.save()
                                                if contador >= capacidad:
                                                    break
                                    if contador < capacidad:
                                        for eMatricula in eMatriculas:
                                            eMateriaAsignadas = MateriaAsignada.objects.filter(status=True,
                                                                                               matricula=eMatricula).exclude(
                                                Q(pk__in=eMateriaAsignadaPlanificacionSedeVirtualExamenes) | Q(
                                                    materia__asignatura__id=4837))
                                            if eMateriaAsignadas.values("id").exists():
                                                eMateriaAsignada = eMateriaAsignadas.first()
                                                if not MateriaAsignadaPlanificacionSedeVirtualExamen.objects.filter(
                                                        status=True,
                                                        materiaasignada=eMateriaAsignada,
                                                        aulaplanificacion__turnoplanificacion__horainicio=eTurnoPlanificacionSedeVirtualExamen.horainicio,
                                                        aulaplanificacion__turnoplanificacion__fechaplanificacion=eFechaPlanificacionSedeVirtualExamen,
                                                        detallemodeloevaluativo_id=detallemodeloevaluativo_id).exists():

                                                    contador += 1
                                                    c += 1
                                                    print(
                                                        f"-------({c}) ->  ({contador}) Se asignada el estudiante {eMateriaAsignada.matricula.inscripcion.persona} en la asignatura {eMateriaAsignada.materia.asignatura.nombre}")
                                                    eMateriaAsignadaPlanificacionSedeVirtualExamen = MateriaAsignadaPlanificacionSedeVirtualExamen(
                                                        aulaplanificacion=eAulaPlanificacionSedeVirtualExamen,
                                                        materiaasignada=eMateriaAsignada,
                                                        detallemodeloevaluativo_id=detallemodeloevaluativo_id)
                                                    eMateriaAsignadaPlanificacionSedeVirtualExamen.save()
                                                    if contador >= capacidad:
                                                        break
                                            deletall = MateriaAsignadaPlanificacionSedeVirtualExamen.objects.filter(
                                                materiaasignada__matricula=eMatricula).values('id')
                                            tmateriasall = MateriaAsignada.objects.filter(status=True,
                                                                                          matricula=eMatricula).values(
                                                'id')
                                            if len(deletall) == len(tmateriasall) and not eMatricula.id in exlall:
                                                exlall.append(eMatricula.id)
                                else:
                                    comp.remove(fmatricula)
                            if not comp:
                                for eMatricula in eMatriculas:
                                    if not eMatricula.id in exlall:
                                        eMateriaAsignadas = MateriaAsignada.objects.filter(status=True,
                                                                                           matricula=eMatricula).exclude(
                                            Q(pk__in=eMateriaAsignadaPlanificacionSedeVirtualExamenes) | Q(
                                                materia__asignatura__id=4837))
                                        if eMateriaAsignadas.values("id").exists():
                                            eMateriaAsignada = eMateriaAsignadas.first()
                                            contador += 1
                                            c += 1
                                            print(
                                                f"-------({c}) -> ({contador}) Se asignada el estudiante {eMateriaAsignada.matricula.inscripcion.persona} en la asignatura {eMateriaAsignada.materia.asignatura.nombre}")
                                            eMateriaAsignadaPlanificacionSedeVirtualExamen = MateriaAsignadaPlanificacionSedeVirtualExamen(
                                                aulaplanificacion=eAulaPlanificacionSedeVirtualExamen,
                                                materiaasignada=eMateriaAsignada,
                                                detallemodeloevaluativo_id=detallemodeloevaluativo_id)
                                            eMateriaAsignadaPlanificacionSedeVirtualExamen.save()
                                            if contador >= capacidad:
                                                break
                                        deletall = MateriaAsignadaPlanificacionSedeVirtualExamen.objects.filter(
                                            status=True,
                                            materiaasignada__matricula_id=eMatricula.id).exclude(
                                            Q(materiaasignada__materia__asignatura__id=4837) |
                                            Q(materiaasignada__materia__asignaturamalla__malla_id__in=eMallasIngles.values_list(
                                                "id", flat=True))).values('id')
                                        tmateriasall = MateriaAsignada.objects.filter(status=True,
                                                                                      matricula=eMatricula).values('id')
                                        if len(deletall) == len(tmateriasall) and not eMatricula.id in exlall:
                                            exlall.append(eMatricula.id)
                comp.extend(list(MateriaAsignadaPlanificacionSedeVirtualExamen.objects.filter(
                    aulaplanificacion__turnoplanificacion__fechaplanificacion__fecha=eFechaPlanificacionSedeVirtualExamen.fecha).values(
                    'aulaplanificacion__turnoplanificacion__fechaplanificacion__fecha').annotate(
                    total=Count('aulaplanificacion__turnoplanificacion__fechaplanificacion__fecha')).values(
                    'total').filter(total__lt=limite_materia).values_list('materiaasignada__matricula_id',
                                                                          flat=True).order_by(
                    'materiaasignada__matricula__inscripcion__persona__apellido1',
                    'materiaasignada__matricula__inscripcion__persona__apellido2',
                    'materiaasignada__matricula__inscripcion__persona__nombres')))
    except Exception as e:
        print(e)
        print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))


# PARA CREAR PLANIFICACION DE FECHAS AULAS Y TURNOS EN CADA SEDE, EN CADA FUNCION EXISTE LOS TURNOS, LAS FECHAS SON GENERAL PARA LAS DOS SEDES


def crear_planificacion_pregrado_sede_examen_milagro():
    horaslv = [
        [(datetime(2023, 9, 5, 8, 45, 0)).time(), (datetime(2023, 9, 5, 9, 59, 0)).time()],
        [(datetime(2023, 9, 5, 10, 00, 0)).time(), (datetime(2023, 9, 5, 11, 14, 0)).time()],
        [(datetime(2023, 9, 5, 11, 15, 0)).time(), (datetime(2023, 9, 5, 12, 29, 0)).time()],
        [(datetime(2023, 9, 5, 13, 45, 0)).time(), (datetime(2023, 9, 5, 14, 59, 0)).time()],
        [(datetime(2023, 9, 5, 15, 00, 0)).time(), (datetime(2023, 9, 5, 16, 14, 0)).time()],
        [(datetime(2023, 9, 5, 16, 15, 0)).time(), (datetime(2023, 9, 5, 17, 29, 0)).time()],
        # [(datetime(2023, 9, 5, 18, 45, 0)).time(), (datetime(2023, 9, 5, 19, 59, 0)).time()],
        # [(datetime(2023, 9, 5, 20, 15, 0)).time(), (datetime(2023, 9, 5, 21, 14, 0)).time()],
        # [(datetime(2023, 9, 5, 14, 15, 0)).time(), (datetime(2023, 9, 5, 15, 29, 0)).time()],
    ]
    sede_id = 1
    for fecha in fechasplan:
        print(f"*** FECHA: {fecha}")
        eFechaPlanificacionSedeVirtualExamenes = FechaPlanificacionSedeVirtualExamen.objects.filter(sede_id=sede_id,
                                                                                                    periodo_id=177,
                                                                                                    fecha=fecha)
        if not eFechaPlanificacionSedeVirtualExamenes.values("id").exists():
            eFechaPlanificacionSedeVirtualExamen = FechaPlanificacionSedeVirtualExamen(sede_id=sede_id,
                                                                                       periodo_id=177,
                                                                                       fecha=fecha)
            eFechaPlanificacionSedeVirtualExamen.save()

        eFechaPlanificacionSedeVirtualExamen = eFechaPlanificacionSedeVirtualExamenes.first()
        # if eFechaPlanificacionSedeVirtualExamen.fecha.weekday() >= 5:
        #     horas = horasfds
        # else:
        horas = horaslv
        for hora in horas:
            horainicio = hora[0]
            horafin = hora[1]
            print(f"*** FECHA: {fecha} -> hora: {horainicio} - {horafin}")
            eTurnoPlanificacionSedeVirtualExamenes = TurnoPlanificacionSedeVirtualExamen.objects.filter(
                fechaplanificacion=eFechaPlanificacionSedeVirtualExamen, horainicio=horainicio, horafin=horafin)
            if not eTurnoPlanificacionSedeVirtualExamenes.values("id").exists():
                eTurnoPlanificacionSedeVirtualExamen = TurnoPlanificacionSedeVirtualExamen(
                    fechaplanificacion=eFechaPlanificacionSedeVirtualExamen,
                    horainicio=horainicio,
                    horafin=horafin)
                eTurnoPlanificacionSedeVirtualExamen.save()
            else:
                eTurnoPlanificacionSedeVirtualExamen = eTurnoPlanificacionSedeVirtualExamenes.first()

            for eLaboratorioVirtual in LaboratorioVirtual.objects.filter(sedevirtual_id=sede_id, activo=True):
                print(f"*** FECHA: {fecha} -> hora: {horainicio} - {horafin} -> aula: {eLaboratorioVirtual.nombre}")
                eAulaPlanificacionSedeVirtualExamenes = AulaPlanificacionSedeVirtualExamen.objects.filter(
                    turnoplanificacion=eTurnoPlanificacionSedeVirtualExamen, aula=eLaboratorioVirtual)
                if not eAulaPlanificacionSedeVirtualExamenes.values("id").exists():
                    eAulaPlanificacionSedeVirtualExamen = AulaPlanificacionSedeVirtualExamen(
                        turnoplanificacion=eTurnoPlanificacionSedeVirtualExamen,
                        aula=eLaboratorioVirtual)
                    eAulaPlanificacionSedeVirtualExamen.save()
                # else:
                #     eAulaPlanificacionSedeVirtualExamen = eAulaPlanificacionSedeVirtualExamenes.first()
    print(f"********************************FINALIZA PROCESO PLANIFICAR SEDE MILAGRO")
    print("")
    print("")
    print("")
    print("")
    print("")


def crear_planificacion_pregrado_sede_examen_santo_domingo():
    horaslv = [
        [(datetime(2023, 9, 5, 7, 30, 0)).time(), (datetime(2023, 9, 5, 8, 44, 0)).time()],
        [(datetime(2023, 9, 5, 8, 45, 0)).time(), (datetime(2023, 9, 5, 9, 59, 0)).time()],
        [(datetime(2023, 9, 5, 10, 00, 0)).time(), (datetime(2023, 9, 5, 11, 14, 0)).time()],
        [(datetime(2023, 9, 5, 11, 15, 0)).time(), (datetime(2023, 9, 5, 12, 29, 0)).time()],
        [(datetime(2023, 9, 5, 13, 45, 0)).time(), (datetime(2023, 9, 5, 14, 59, 0)).time()],
        [(datetime(2023, 9, 5, 15, 00, 0)).time(), (datetime(2023, 9, 5, 16, 14, 0)).time()],
        [(datetime(2023, 9, 5, 16, 15, 0)).time(), (datetime(2023, 9, 5, 17, 29, 0)).time()],
        [(datetime(2023, 9, 5, 17, 30, 0)).time(), (datetime(2023, 9, 5, 18, 44, 0)).time()],
    ]
    sede_id = 10
    for fecha in fechasplan:
        print(f"*** FECHA: {fecha}")
        eFechaPlanificacionSedeVirtualExamenes = FechaPlanificacionSedeVirtualExamen.objects.filter(sede_id=sede_id,
                                                                                                    periodo_id=177,
                                                                                                    fecha=fecha)
        if not eFechaPlanificacionSedeVirtualExamenes.values("id").exists():
            eFechaPlanificacionSedeVirtualExamen = FechaPlanificacionSedeVirtualExamen(sede_id=sede_id,
                                                                                       periodo_id=177,
                                                                                       fecha=fecha)
            eFechaPlanificacionSedeVirtualExamen.save()

        eFechaPlanificacionSedeVirtualExamen = eFechaPlanificacionSedeVirtualExamenes.first()
        # if eFechaPlanificacionSedeVirtualExamen.fecha.weekday() >= 5:
        #     horas = horasfds
        # else:
        horas = horaslv
        for hora in horas:
            horainicio = hora[0]
            horafin = hora[1]
            print(f"*** FECHA: {fecha} -> hora: {horainicio} - {horafin}")
            eTurnoPlanificacionSedeVirtualExamenes = TurnoPlanificacionSedeVirtualExamen.objects.filter(
                fechaplanificacion=eFechaPlanificacionSedeVirtualExamen, horainicio=horainicio, horafin=horafin)
            if not eTurnoPlanificacionSedeVirtualExamenes.values("id").exists():
                eTurnoPlanificacionSedeVirtualExamen = TurnoPlanificacionSedeVirtualExamen(
                    fechaplanificacion=eFechaPlanificacionSedeVirtualExamen,
                    horainicio=horainicio,
                    horafin=horafin)
                eTurnoPlanificacionSedeVirtualExamen.save()
            else:
                eTurnoPlanificacionSedeVirtualExamen = eTurnoPlanificacionSedeVirtualExamenes.first()

            for eLaboratorioVirtual in LaboratorioVirtual.objects.filter(sedevirtual_id=sede_id, activo=True):
                print(f"*** FECHA: {fecha} -> hora: {horainicio} - {horafin} -> aula: {eLaboratorioVirtual.nombre}")
                eAulaPlanificacionSedeVirtualExamenes = AulaPlanificacionSedeVirtualExamen.objects.filter(
                    turnoplanificacion=eTurnoPlanificacionSedeVirtualExamen, aula=eLaboratorioVirtual)
                if not eAulaPlanificacionSedeVirtualExamenes.values("id").exists():
                    eAulaPlanificacionSedeVirtualExamen = AulaPlanificacionSedeVirtualExamen(
                        turnoplanificacion=eTurnoPlanificacionSedeVirtualExamen,
                        aula=eLaboratorioVirtual)
                    eAulaPlanificacionSedeVirtualExamen.save()
                # else:
                #     eAulaPlanificacionSedeVirtualExamen = eAulaPlanificacionSedeVirtualExamenes.first()
    print(f"********************************FINALIZA PROCESO PLANIFICAR SEDE SANTO DOMINGO")
    print("")
    print("")
    print("")
    print("")
    print("")
    crear_planificacion_pregrado_sede_examen_milagro()


def crear_planificacion_pregrado_sede_examen_virtual():
    horaslv = [
        [(datetime(2023, 9, 5, 7, 30, 0)).time(), (datetime(2023, 9, 5, 16, 59, 0)).time()]
    ]
    sede_id = 11
    for fecha in fechasplan:
        print(f"*** FECHA: {fecha}")
        eFechaPlanificacionSedeVirtualExamenes = FechaPlanificacionSedeVirtualExamen.objects.filter(sede_id=sede_id,
                                                                                                    periodo_id=177,
                                                                                                    fecha=fecha)
        if not eFechaPlanificacionSedeVirtualExamenes.values("id").exists():
            eFechaPlanificacionSedeVirtualExamen = FechaPlanificacionSedeVirtualExamen(sede_id=sede_id,
                                                                                       periodo_id=177,
                                                                                       fecha=fecha)
            eFechaPlanificacionSedeVirtualExamen.save()

        eFechaPlanificacionSedeVirtualExamen = eFechaPlanificacionSedeVirtualExamenes.first()
        # if eFechaPlanificacionSedeVirtualExamen.fecha.weekday() >= 5:
        #     horas = horasfds
        # else:
        horas = horaslv
        for hora in horas:
            horainicio = hora[0]
            horafin = hora[1]
            print(f"*** FECHA: {fecha} -> hora: {horainicio} - {horafin}")
            eTurnoPlanificacionSedeVirtualExamenes = TurnoPlanificacionSedeVirtualExamen.objects.filter(
                fechaplanificacion=eFechaPlanificacionSedeVirtualExamen, horainicio=horainicio, horafin=horafin)
            if not eTurnoPlanificacionSedeVirtualExamenes.values("id").exists():
                eTurnoPlanificacionSedeVirtualExamen = TurnoPlanificacionSedeVirtualExamen(
                    fechaplanificacion=eFechaPlanificacionSedeVirtualExamen,
                    horainicio=horainicio,
                    horafin=horafin)
                eTurnoPlanificacionSedeVirtualExamen.save()
            else:
                eTurnoPlanificacionSedeVirtualExamen = eTurnoPlanificacionSedeVirtualExamenes.first()

            for eLaboratorioVirtual in LaboratorioVirtual.objects.filter(sedevirtual_id=sede_id, activo=True):
                print(f"*** FECHA: {fecha} -> hora: {horainicio} - {horafin} -> aula: {eLaboratorioVirtual.nombre}")
                eAulaPlanificacionSedeVirtualExamenes = AulaPlanificacionSedeVirtualExamen.objects.filter(
                    turnoplanificacion=eTurnoPlanificacionSedeVirtualExamen, aula=eLaboratorioVirtual)
                if not eAulaPlanificacionSedeVirtualExamenes.values("id").exists():
                    eAulaPlanificacionSedeVirtualExamen = AulaPlanificacionSedeVirtualExamen(
                        turnoplanificacion=eTurnoPlanificacionSedeVirtualExamen,
                        aula=eLaboratorioVirtual)
                    eAulaPlanificacionSedeVirtualExamen.save()
                # else:
                #     eAulaPlanificacionSedeVirtualExamen = eAulaPlanificacionSedeVirtualExamenes.first()
    print(f"********************************FINALIZA PROCESO PLANIFICAR SEDE VIRTUAL")
    print("")
    print("")
    print("")
    print("")
    print("")
    crear_planificacion_pregrado_sede_examen_santo_domingo()


def crear_planificacion_nivelacion_sede_examen_milagro():
    horaslv = [
        [(datetime(2023, 9, 5, 8, 00, 0)).time(), (datetime(2023, 9, 5, 9, 14, 0)).time()],
        [(datetime(2023, 9, 5, 9, 15, 0)).time(), (datetime(2023, 9, 5, 10, 29, 0)).time()],
        [(datetime(2023, 9, 5, 10, 30, 0)).time(), (datetime(2023, 9, 5, 11, 44, 0)).time()],
        [(datetime(2023, 9, 5, 11, 45, 0)).time(), (datetime(2023, 9, 5, 12, 59, 0)).time()],
        [(datetime(2023, 9, 5, 13, 00, 0)).time(), (datetime(2023, 9, 5, 14, 14, 0)).time()],
        [(datetime(2023, 9, 5, 15, 00, 0)).time(), (datetime(2023, 9, 5, 16, 14, 0)).time()],
        [(datetime(2023, 9, 5, 16, 15, 0)).time(), (datetime(2023, 9, 5, 17, 29, 0)).time()],
        [(datetime(2023, 9, 5, 17, 30, 0)).time(), (datetime(2023, 9, 5, 18, 44, 0)).time()],
        [(datetime(2023, 9, 5, 18, 45, 0)).time(), (datetime(2023, 9, 5, 19, 59, 0)).time()],
        [(datetime(2023, 9, 5, 20, 45, 0)).time(), (datetime(2023, 9, 5, 21, 14, 0)).time()],
    ]
    sede_id = 1
    for fecha in fechasplan_nivelacion:
        print(f"*** FECHA: {fecha}")
        eFechaPlanificacionSedeVirtualExamenes = FechaPlanificacionSedeVirtualExamen.objects.filter(sede_id=sede_id,
                                                                                                    periodo_id=177,
                                                                                                    fecha=fecha)
        if not eFechaPlanificacionSedeVirtualExamenes.values("id").exists():
            eFechaPlanificacionSedeVirtualExamen = FechaPlanificacionSedeVirtualExamen(sede_id=sede_id,
                                                                                       periodo_id=177,
                                                                                       fecha=fecha)
            eFechaPlanificacionSedeVirtualExamen.save()

        eFechaPlanificacionSedeVirtualExamen = eFechaPlanificacionSedeVirtualExamenes.first()
        # if eFechaPlanificacionSedeVirtualExamen.fecha.weekday() >= 5:
        #     horas = horasfds
        # else:
        horas = horaslv
        for hora in horas:
            horainicio = hora[0]
            horafin = hora[1]
            print(f"*** FECHA: {fecha} -> hora: {horainicio} - {horafin}")
            eTurnoPlanificacionSedeVirtualExamenes = TurnoPlanificacionSedeVirtualExamen.objects.filter(
                fechaplanificacion=eFechaPlanificacionSedeVirtualExamen, horainicio=horainicio, horafin=horafin)
            if not eTurnoPlanificacionSedeVirtualExamenes.values("id").exists():
                eTurnoPlanificacionSedeVirtualExamen = TurnoPlanificacionSedeVirtualExamen(
                    fechaplanificacion=eFechaPlanificacionSedeVirtualExamen,
                    horainicio=horainicio,
                    horafin=horafin)
                eTurnoPlanificacionSedeVirtualExamen.save()
            else:
                eTurnoPlanificacionSedeVirtualExamen = eTurnoPlanificacionSedeVirtualExamenes.first()

            for eLaboratorioVirtual in LaboratorioVirtual.objects.filter(sedevirtual_id=sede_id, activo=True):
                print(f"*** FECHA: {fecha} -> hora: {horainicio} - {horafin} -> aula: {eLaboratorioVirtual.nombre}")
                eAulaPlanificacionSedeVirtualExamenes = AulaPlanificacionSedeVirtualExamen.objects.filter(
                    turnoplanificacion=eTurnoPlanificacionSedeVirtualExamen, aula=eLaboratorioVirtual)
                if not eAulaPlanificacionSedeVirtualExamenes.values("id").exists():
                    eAulaPlanificacionSedeVirtualExamen = AulaPlanificacionSedeVirtualExamen(
                        turnoplanificacion=eTurnoPlanificacionSedeVirtualExamen,
                        aula=eLaboratorioVirtual)
                    eAulaPlanificacionSedeVirtualExamen.save()
                # else:
                #     eAulaPlanificacionSedeVirtualExamen = eAulaPlanificacionSedeVirtualExamenes.first()
    print(f"********************************FINALIZA PROCESO PLANIFICAR SEDE VIRTUAL")
    print("")
    print("")
    print("")
    print("")
    print("")
    crear_planificacion_pregrado_sede_examen_virtual()


def crear_planificacion_nivelacion_sede_examen_santo_domingo():
    horaslv = [
        [(datetime(2023, 9, 5, 8, 00, 0)).time(), (datetime(2023, 9, 5, 9, 14, 0)).time()],
        [(datetime(2023, 9, 5, 9, 15, 0)).time(), (datetime(2023, 9, 5, 10, 29, 0)).time()],
        [(datetime(2023, 9, 5, 10, 30, 0)).time(), (datetime(2023, 9, 5, 11, 44, 0)).time()],
        [(datetime(2023, 9, 5, 11, 45, 0)).time(), (datetime(2023, 9, 5, 12, 59, 0)).time()],
        [(datetime(2023, 9, 5, 13, 00, 0)).time(), (datetime(2023, 9, 5, 14, 14, 0)).time()],
        [(datetime(2023, 9, 5, 15, 00, 0)).time(), (datetime(2023, 9, 5, 16, 14, 0)).time()],
        [(datetime(2023, 9, 5, 16, 15, 0)).time(), (datetime(2023, 9, 5, 17, 29, 0)).time()],
        [(datetime(2023, 9, 5, 17, 30, 0)).time(), (datetime(2023, 9, 5, 18, 44, 0)).time()],
        [(datetime(2023, 9, 5, 18, 45, 0)).time(), (datetime(2023, 9, 5, 19, 59, 0)).time()],
        [(datetime(2023, 9, 5, 20, 45, 0)).time(), (datetime(2023, 9, 5, 21, 14, 0)).time()],
    ]
    sede_id = 10
    for fecha in fechasplan_nivelacion:
        print(f"*** FECHA: {fecha}")
        eFechaPlanificacionSedeVirtualExamenes = FechaPlanificacionSedeVirtualExamen.objects.filter(sede_id=sede_id,
                                                                                                    periodo_id=177,
                                                                                                    fecha=fecha)
        if not eFechaPlanificacionSedeVirtualExamenes.values("id").exists():
            eFechaPlanificacionSedeVirtualExamen = FechaPlanificacionSedeVirtualExamen(sede_id=sede_id,
                                                                                       periodo_id=177,
                                                                                       fecha=fecha)
            eFechaPlanificacionSedeVirtualExamen.save()

        eFechaPlanificacionSedeVirtualExamen = eFechaPlanificacionSedeVirtualExamenes.first()
        # if eFechaPlanificacionSedeVirtualExamen.fecha.weekday() >= 5:
        #     horas = horasfds
        # else:
        horas = horaslv
        for hora in horas:
            horainicio = hora[0]
            horafin = hora[1]
            print(f"*** FECHA: {fecha} -> hora: {horainicio} - {horafin}")
            eTurnoPlanificacionSedeVirtualExamenes = TurnoPlanificacionSedeVirtualExamen.objects.filter(
                fechaplanificacion=eFechaPlanificacionSedeVirtualExamen, horainicio=horainicio, horafin=horafin)
            if not eTurnoPlanificacionSedeVirtualExamenes.values("id").exists():
                eTurnoPlanificacionSedeVirtualExamen = TurnoPlanificacionSedeVirtualExamen(
                    fechaplanificacion=eFechaPlanificacionSedeVirtualExamen,
                    horainicio=horainicio,
                    horafin=horafin)
                eTurnoPlanificacionSedeVirtualExamen.save()
            else:
                eTurnoPlanificacionSedeVirtualExamen = eTurnoPlanificacionSedeVirtualExamenes.first()

            for eLaboratorioVirtual in LaboratorioVirtual.objects.filter(sedevirtual_id=sede_id, activo=True):
                print(f"*** FECHA: {fecha} -> hora: {horainicio} - {horafin} -> aula: {eLaboratorioVirtual.nombre}")
                eAulaPlanificacionSedeVirtualExamenes = AulaPlanificacionSedeVirtualExamen.objects.filter(
                    turnoplanificacion=eTurnoPlanificacionSedeVirtualExamen, aula=eLaboratorioVirtual)
                if not eAulaPlanificacionSedeVirtualExamenes.values("id").exists():
                    eAulaPlanificacionSedeVirtualExamen = AulaPlanificacionSedeVirtualExamen(
                        turnoplanificacion=eTurnoPlanificacionSedeVirtualExamen,
                        aula=eLaboratorioVirtual)
                    eAulaPlanificacionSedeVirtualExamen.save()
                # else:
                #     eAulaPlanificacionSedeVirtualExamen = eAulaPlanificacionSedeVirtualExamenes.first()
    print(f"********************************FINALIZA PROCESO PLANIFICAR SEDE SANTO DOMINGO")
    print("")
    print("")
    print("")
    print("")
    print("")
    crear_planificacion_nivelacion_sede_examen_milagro()


def crear_planificacion_nivelacion_sede_examen_virtual():
    horaslv = [
        [(datetime(2023, 9, 5, 7, 30, 0)).time(), (datetime(2023, 9, 5, 16, 59, 0)).time()]
    ]
    sede_id = 11
    for fecha in fechasplan_nivelacion:
        print(f"*** FECHA: {fecha}")
        eFechaPlanificacionSedeVirtualExamenes = FechaPlanificacionSedeVirtualExamen.objects.filter(sede_id=sede_id,
                                                                                                    periodo_id=177,
                                                                                                    fecha=fecha)
        if not eFechaPlanificacionSedeVirtualExamenes.values("id").exists():
            eFechaPlanificacionSedeVirtualExamen = FechaPlanificacionSedeVirtualExamen(sede_id=sede_id,
                                                                                       periodo_id=177,
                                                                                       fecha=fecha)
            eFechaPlanificacionSedeVirtualExamen.save()

        eFechaPlanificacionSedeVirtualExamen = eFechaPlanificacionSedeVirtualExamenes.first()
        # if eFechaPlanificacionSedeVirtualExamen.fecha.weekday() >= 5:
        #     horas = horasfds
        # else:
        horas = horaslv
        for hora in horas:
            horainicio = hora[0]
            horafin = hora[1]
            print(f"*** FECHA: {fecha} -> hora: {horainicio} - {horafin}")
            eTurnoPlanificacionSedeVirtualExamenes = TurnoPlanificacionSedeVirtualExamen.objects.filter(
                fechaplanificacion=eFechaPlanificacionSedeVirtualExamen, horainicio=horainicio, horafin=horafin)
            if not eTurnoPlanificacionSedeVirtualExamenes.values("id").exists():
                eTurnoPlanificacionSedeVirtualExamen = TurnoPlanificacionSedeVirtualExamen(
                    fechaplanificacion=eFechaPlanificacionSedeVirtualExamen,
                    horainicio=horainicio,
                    horafin=horafin)
                eTurnoPlanificacionSedeVirtualExamen.save()
            else:
                eTurnoPlanificacionSedeVirtualExamen = eTurnoPlanificacionSedeVirtualExamenes.first()

            for eLaboratorioVirtual in LaboratorioVirtual.objects.filter(sedevirtual_id=sede_id, activo=True):
                print(f"*** FECHA: {fecha} -> hora: {horainicio} - {horafin} -> aula: {eLaboratorioVirtual.nombre}")
                eAulaPlanificacionSedeVirtualExamenes = AulaPlanificacionSedeVirtualExamen.objects.filter(
                    turnoplanificacion=eTurnoPlanificacionSedeVirtualExamen, aula=eLaboratorioVirtual)
                if not eAulaPlanificacionSedeVirtualExamenes.values("id").exists():
                    eAulaPlanificacionSedeVirtualExamen = AulaPlanificacionSedeVirtualExamen(
                        turnoplanificacion=eTurnoPlanificacionSedeVirtualExamen,
                        aula=eLaboratorioVirtual)
                    eAulaPlanificacionSedeVirtualExamen.save()
                # else:
                #     eAulaPlanificacionSedeVirtualExamen = eAulaPlanificacionSedeVirtualExamenes.first()
    print(f"********************************FINALIZA PROCESO PLANIFICAR SEDE VIRTUAL")
    print("")
    print("")
    print("")
    print("")
    print("")
    crear_planificacion_nivelacion_sede_examen_santo_domingo()


# crear_planificacion_nivelacion_sede_examen_virtual()


# CREAR LOS HORARIOS DE LAS MATERIAS A MOSTRAR

def save_data(c, materia, aula, planificado_materias_aula, detallemodeloevaluativo_id):
    print('N° {} -- Nivel: {} -- Materia: {} --- Aula: {} -- Cupo: {}'.format(c,
                                                                              materia.asignaturamalla.nivelmalla.id,
                                                                              materia.asignatura.nombre + ' ' + materia.paralelo,
                                                                              str(aula.aula) + ' ' +
                                                                              str(aula.turnoplanificacion.fechaplanificacion.fecha) + ' -- ' +
                                                                              str(aula.turnoplanificacion.horainicio) + ' hasta ' +
                                                                              str(aula.turnoplanificacion.horafin),
                                                                              planificado_materias_aula))
    if not MateriaPlanificacionSedeVirtualExamen.objects.filter(materia=materia,
                                                                aulaplanificacion=aula,
                                                                detallemodeloevaluativo_id=detallemodeloevaluativo_id,
                                                                asignaturamalla=materia.asignaturamalla).exists():
        planmateria = MateriaPlanificacionSedeVirtualExamen(materia=materia,
                                                            aulaplanificacion=aula,
                                                            detallemodeloevaluativo_id=detallemodeloevaluativo_id,
                                                            asignaturamalla=materia.asignaturamalla,
                                                            cupo=planificado_materias_aula)
        planmateria.save()


def crear_plan_materias(sede_id, periodo, fechas):
    try:
        set_ppl_statatus()
        # crear_planificacion_pregrado_sede_examen_milagro()
        mat = []
        exl = []
        detallemodeloevaluativo_id = 37
        eMallasIngles = Malla.objects.filter(pk__in=[353, 22]).values_list('id', flat=True)
        eMallas = Malla.objects.filter(carrera__modalidad=3, carrera__coordinacion__lte=5, carrera_id=127).exclude(
            id__in=eMallasIngles)
        eMallas = eMallas.order_by('carrera__nombre', 'inicio')
        materia = None
        asignaturamalla = None
        planificado_materias = 0
        totalasignaturas = 0
        matriculas = MatriculaSedeExamen.objects.filter(status=True, sede_id=sede_id,
                                                        detallemodeloevaluativo_id=detallemodeloevaluativo_id,
                                                        matricula__nivel__periodo_id=periodo).values_list(
            'matricula_id', flat=True)

        c = 0
        for eMalla in eMallas:
            if not Materia.objects.filter(status=True, nivel__periodo_id=177, asignaturamalla__malla=eMalla).exists():
                continue
            for fecha in FechaPlanificacionSedeVirtualExamen.objects.filter(sede_id=sede_id, periodo_id=periodo,
                                                                            fecha__in=fechas):
                matasig = []
                niveles = []
                totalasignaturas = 0
                for turno in fecha.turnoplanificacionsedevirtualexamen_set.filter(status=True):
                    for aula in turno.aulaplanificacionsedevirtualexamen_set.filter(status=True).exclude(id__in=exl):
                        capacidad = aula.aula.capacidad
                        cantidadad_planificadas = aula.cantidadad_planificadas_materias()
                        if cantidadad_planificadas >= capacidad:
                            break
                        planificado_materias_aula = 0
                        materias = MateriaAsignada.objects.filter(status=True, materia__nivel__periodo_id=177,
                                                                  retiramateria=False,
                                                                  matricula__retiradomatricula=False,
                                                                  matricula__inscripcion__carrera__modalidad=3,
                                                                  materia__asignaturamalla__malla__carrera__coordinacion__lte=5,
                                                                  materia__asignaturamalla__malla=eMalla).filter(
                            matricula_id__in=matriculas).exclude(
                            materia_id__in=mat).exclude(materia__asignaturamalla__malla_id__in=eMallasIngles).exclude(
                            id__in=matasig).exclude(materia__asignaturamalla__nivelmalla_id__in=niveles).exclude(
                            matricula__inscripcion__persona__ppl=True).exclude(
                            matricula__inscripcion__persona__perfilinscripcion__tienediscapacidad=True).order_by(
                            'materia__asignaturamalla__nivelmalla_id', 'materia__asignatura__nombre',
                            'materia__paralelo')
                        if not materias:
                            break
                        for materiaasig in materias:
                            nivel = materiaasig.materia.asignaturamalla.nivelmalla_id
                            if totalasignaturas == 4 and nivel not in niveles:
                                if len(MateriaPlanificacionSedeVirtualExamen.objects.filter(status=True,
                                                                                            asignaturamalla__nivelmalla_id=nivel,
                                                                                            asignaturamalla__malla=eMalla,
                                                                                            aulaplanificacion__turnoplanificacion__fechaplanificacion__fecha=fecha.fecha).values(
                                    'asignaturamalla__nivelmalla_id',
                                    'aulaplanificacion__turnoplanificacion__fechaplanificacion__fecha',
                                    'asignaturamalla_id').distinct().order_by(
                                    'aulaplanificacion__turnoplanificacion__fechaplanificacion__fecha',
                                    'asignaturamalla_id')) == 3:
                                    niveles.append(materiaasig.materia.asignaturamalla.nivelmalla_id)
                                else:
                                    totalasignaturas = 1
                            if materiaasig.materia.asignaturamalla.nivelmalla_id in niveles:
                                totalasignaturas = 0
                                continue
                            if not materia == materiaasig.materia:
                                planificado_materias = 0
                                planificado_materias_aula = 0
                            materia = materiaasig.materia

                            materia_id = materia.id
                            totalmateria = len(
                                materia.asignados_a_esta_materia_todos().filter(matricula_id__in=matriculas).exclude(
                                    matricula__inscripcion__persona__ppl=True).exclude(
                                    matricula__inscripcion__persona__perfilinscripcion__tienediscapacidad=True))
                            # totalmateria = int(totalmateria - (totalmateria * 0.09))
                            # cantidadad_planificadas = aula.cantidadad_planificadas()
                            if not asignaturamalla == materiaasig.materia.asignaturamalla:
                                totalasignaturas += 1
                            asignaturamalla = materiaasig.materia.asignaturamalla
                            if cantidadad_planificadas < capacidad:
                                if materia_id not in mat:
                                    c += 1
                                    matasig.append(materiaasig.pk)
                                    cantidadad_planificadas += 1
                                    planificado_materias += 1
                                    planificado_materias_aula += 1
                                    if planificado_materias == totalmateria and materia_id not in mat:
                                        mat.append(materiaasig.materia_id)
                                        save_data(c, materia, aula, planificado_materias_aula,
                                                  detallemodeloevaluativo_id)
                                        continue
                            else:
                                if materia_id not in mat and planificado_materias_aula > 1:
                                    save_data(c, materia, aula, planificado_materias_aula, detallemodeloevaluativo_id)
                                break

    except Exception as ex:
        print(ex)


def crear_plan_materias_sto_domingo(sede_id, periodo, fechas):
    try:
        set_ppl_statatus()
        mat = []
        exl = []
        detallemodeloevaluativo_id = 37
        eMallasIngles = Malla.objects.filter(pk__in=[353, 22]).values_list('id', flat=True)
        eMallas = Malla.objects.filter(carrera__modalidad=3, carrera__coordinacion__lte=5).exclude(id__in=eMallasIngles)
        eMallas = eMallas.order_by('carrera__nombre', 'inicio')
        materia = None
        asignaturamalla = None
        planificado_materias = 0
        totalasignaturas = 0

        c = 0
        for eMalla in eMallas:
            if not Materia.objects.filter(status=True, nivel__periodo_id=177, asignaturamalla__malla=eMalla).exists():
                continue
            for fecha in FechaPlanificacionSedeVirtualExamen.objects.filter(sede_id=sede_id, periodo_id=periodo,
                                                                            fecha__in=fechas):
                matasig = []
                niveles = []
                totalasignaturas = 0
                for turno in fecha.turnoplanificacionsedevirtualexamen_set.filter(status=True):
                    for aula in turno.aulaplanificacionsedevirtualexamen_set.filter(status=True).exclude(id__in=exl):
                        capacidad = aula.aula.capacidad
                        cantidadad_planificadas = aula.cantidadad_planificadas_materias()
                        if cantidadad_planificadas >= capacidad:
                            break
                        planificado_materias_aula = 0
                        materias = MateriaAsignada.objects.filter(status=True, materia__nivel__periodo_id=177,
                                                                  retiramateria=False,
                                                                  matricula__retiradomatricula=False,
                                                                  matricula__inscripcion__carrera__modalidad=3,
                                                                  materia__asignaturamalla__malla__carrera__coordinacion__lte=5,
                                                                  materia__asignaturamalla__malla=eMalla).exclude(
                            materia_id__in=mat).exclude(materia__asignaturamalla__malla_id__in=eMallasIngles).exclude(
                            id__in=matasig).exclude(materia__asignaturamalla__nivelmalla_id__in=niveles).exclude(
                            matricula__inscripcion__persona__ppl=True).exclude(
                            matricula__inscripcion__persona__perfilinscripcion__tienediscapacidad=True).order_by(
                            'materia__asignaturamalla__nivelmalla_id', 'materia__asignatura__nombre',
                            'materia__paralelo')
                        if not materias:
                            break
                        for materiaasig in materias:
                            nivel = materiaasig.materia.asignaturamalla.nivelmalla_id
                            if totalasignaturas == 4 and nivel not in niveles:
                                if len(MateriaPlanificacionSedeVirtualExamen.objects.filter(status=True,
                                                                                            asignaturamalla__nivelmalla_id=nivel,
                                                                                            asignaturamalla__malla=eMalla,
                                                                                            aulaplanificacion__turnoplanificacion__fechaplanificacion__fecha=fecha.fecha).values(
                                    'asignaturamalla__nivelmalla_id',
                                    'aulaplanificacion__turnoplanificacion__fechaplanificacion__fecha',
                                    'asignaturamalla_id').distinct().order_by(
                                    'aulaplanificacion__turnoplanificacion__fechaplanificacion__fecha',
                                    'asignaturamalla_id')) == 3:
                                    niveles.append(materiaasig.materia.asignaturamalla.nivelmalla_id)
                                else:
                                    totalasignaturas = 1
                            if materiaasig.materia.asignaturamalla.nivelmalla_id in niveles:
                                totalasignaturas = 0
                                continue
                            if not materia == materiaasig.materia:
                                planificado_materias = 0
                                planificado_materias_aula = 0
                            materia = materiaasig.materia

                            materia_id = materia.id
                            totalmateria = len(materia.asignados_a_esta_materia_todos().exclude(
                                matricula__inscripcion__persona__ppl=True).exclude(
                                matricula__inscripcion__persona__perfilinscripcion__tienediscapacidad=True))
                            totalmateria = int(totalmateria * 0.09)
                            # cantidadad_planificadas = aula.cantidadad_planificadas()
                            if not asignaturamalla == materiaasig.materia.asignaturamalla:
                                totalasignaturas += 1
                            asignaturamalla = materiaasig.materia.asignaturamalla
                            if cantidadad_planificadas < capacidad:
                                if materia_id not in mat:
                                    c += 1
                                    matasig.append(materiaasig.pk)
                                    cantidadad_planificadas += 1
                                    planificado_materias += 1
                                    planificado_materias_aula += 1
                                    if planificado_materias == totalmateria and materia_id not in mat:
                                        mat.append(materiaasig.materia_id)
                                        save_data(c, materia, aula, planificado_materias_aula,
                                                  detallemodeloevaluativo_id)
                                        continue
                            else:
                                if materia_id not in mat and planificado_materias_aula > 1:
                                    save_data(c, materia, aula, planificado_materias_aula, detallemodeloevaluativo_id)
                                break

    except Exception as ex:
        print(ex)


# PLAN DE MATERIAS DE NIVELACION
def crear_plan_materias_nivelacion(sede_id, periodo, fechas):
    try:
        mat = []
        exl = []
        detallemodeloevaluativo_id = 114
        eMallasIngles = Malla.objects.filter(pk__in=[353, 22]).values_list('id', flat=True)
        eMallas = Malla.objects.filter(carrera__modalidad=3, carrera__coordinacion=9, ).exclude(id__in=eMallasIngles)
        eMallas = eMallas.order_by('carrera__nombre', 'inicio')
        materia = None
        asignaturamalla = None
        planificado_materias = 0
        totalasignaturas = 0
        matriculas = MatriculaSedeExamen.objects.filter(status=True, sede_id=sede_id,
                                                        detallemodeloevaluativo_id=detallemodeloevaluativo_id,
                                                        matricula__nivel__periodo_id=periodo).values_list(
            'matricula_id', flat=True)

        c = 0
        for eMalla in eMallas:
            if not Materia.objects.filter(status=True, nivel__periodo_id=177, asignaturamalla__malla=eMalla).exists():
                continue
            for fecha in FechaPlanificacionSedeVirtualExamen.objects.filter(sede_id=sede_id, periodo_id=periodo,
                                                                            fecha__in=fechas):
                matasig = []
                niveles = []
                totalasignaturas = 0
                for turno in fecha.turnoplanificacionsedevirtualexamen_set.filter(status=True):
                    for aula in turno.aulaplanificacionsedevirtualexamen_set.filter(status=True).exclude(id__in=exl):
                        capacidad = aula.aula.capacidad
                        cantidadad_planificadas = aula.cantidadad_planificadas_materias()
                        if cantidadad_planificadas >= capacidad:
                            break
                        planificado_materias_aula = 0
                        materias = MateriaAsignada.objects.filter(status=True, materia__nivel__periodo_id=177,
                                                                  retiramateria=False,
                                                                  matricula__retiradomatricula=False,
                                                                  matricula__inscripcion__carrera__modalidad=3,
                                                                  materia__asignaturamalla__malla__carrera__coordinacion__lte=5,
                                                                  materia__asignaturamalla__malla=eMalla).filter(
                            matricula_id__in=matriculas).exclude(
                            materia_id__in=mat).exclude(materia__asignaturamalla__malla_id__in=eMallasIngles).exclude(
                            id__in=matasig).exclude(materia__asignaturamalla__nivelmalla_id__in=niveles).exclude(
                            matricula__inscripcion__persona__ppl=True).exclude(
                            matricula__inscripcion__persona__perfilinscripcion__tienediscapacidad=True).order_by(
                            'materia__asignaturamalla__nivelmalla_id', 'materia__asignatura__nombre',
                            'materia__paralelo')
                        if not materias:
                            break
                        for materiaasig in materias:
                            nivel = materiaasig.materia.asignaturamalla.nivelmalla_id
                            if totalasignaturas == 4 and nivel not in niveles:
                                if len(MateriaPlanificacionSedeVirtualExamen.objects.filter(status=True,
                                                                                            asignaturamalla__nivelmalla_id=nivel,
                                                                                            asignaturamalla__malla=eMalla,
                                                                                            aulaplanificacion__turnoplanificacion__fechaplanificacion__fecha=fecha.fecha).values(
                                    'asignaturamalla__nivelmalla_id',
                                    'aulaplanificacion__turnoplanificacion__fechaplanificacion__fecha',
                                    'asignaturamalla_id').distinct().order_by(
                                    'aulaplanificacion__turnoplanificacion__fechaplanificacion__fecha',
                                    'asignaturamalla_id')) == 3:
                                    niveles.append(materiaasig.materia.asignaturamalla.nivelmalla_id)
                                else:
                                    totalasignaturas = 1
                            if materiaasig.materia.asignaturamalla.nivelmalla_id in niveles:
                                totalasignaturas = 0
                                continue
                            if not materia == materiaasig.materia:
                                planificado_materias = 0
                                planificado_materias_aula = 0
                            materia = materiaasig.materia

                            materia_id = materia.id
                            totalmateria = len(
                                materia.asignados_a_esta_materia_todos().filter(matricula_id__in=matriculas).exclude(
                                    matricula__inscripcion__persona__ppl=True).exclude(
                                    matricula__inscripcion__persona__perfilinscripcion__tienediscapacidad=True))
                            # totalmateria = int(totalmateria - (totalmateria * 0.09))
                            # cantidadad_planificadas = aula.cantidadad_planificadas()
                            if not asignaturamalla == materiaasig.materia.asignaturamalla:
                                totalasignaturas += 1
                            asignaturamalla = materiaasig.materia.asignaturamalla
                            if cantidadad_planificadas < capacidad:
                                if materia_id not in mat:
                                    c += 1
                                    matasig.append(materiaasig.pk)
                                    cantidadad_planificadas += 1
                                    planificado_materias += 1
                                    planificado_materias_aula += 1
                                    if planificado_materias == totalmateria and materia_id not in mat:
                                        mat.append(materiaasig.materia_id)
                                        save_data(c, materia, aula, planificado_materias_aula,
                                                  detallemodeloevaluativo_id)
                                        continue
                            else:
                                if materia_id not in mat and planificado_materias_aula > 1:
                                    save_data(c, materia, aula, planificado_materias_aula, detallemodeloevaluativo_id)
                                break
    except Exception as ex:
        print(ex)


def crear_plan_materias_nivelacion_sto_domingo(sede_id, periodo, fechas):
    try:
        mat = []
        exl = []
        detallemodeloevaluativo_id = 114
        eMallasIngles = Malla.objects.filter(pk__in=[353, 22]).values_list('id', flat=True)
        eMallas = Malla.objects.filter(carrera__modalidad=3, carrera__coordinacion__lte=5).exclude(id__in=eMallasIngles)
        eMallas = eMallas.order_by('carrera__nombre', 'inicio')
        materia = None
        asignaturamalla = None
        planificado_materias = 0
        totalasignaturas = 0

        c = 0
        for eMalla in eMallas:
            if not Materia.objects.filter(status=True, nivel__periodo_id=177, asignaturamalla__malla=eMalla).exists():
                continue
            for fecha in FechaPlanificacionSedeVirtualExamen.objects.filter(sede_id=sede_id, periodo_id=periodo,
                                                                            fecha__in=fechas):
                matasig = []
                niveles = []
                totalasignaturas = 0
                for turno in fecha.turnoplanificacionsedevirtualexamen_set.filter(status=True):
                    for aula in turno.aulaplanificacionsedevirtualexamen_set.filter(status=True).exclude(id__in=exl):
                        capacidad = aula.aula.capacidad
                        cantidadad_planificadas = aula.cantidadad_planificadas_materias()
                        if cantidadad_planificadas >= capacidad:
                            break
                        planificado_materias_aula = 0
                        materias = MateriaAsignada.objects.filter(status=True, materia__nivel__periodo_id=177,
                                                                  retiramateria=False,
                                                                  matricula__retiradomatricula=False,
                                                                  matricula__inscripcion__carrera__modalidad=3,
                                                                  materia__asignaturamalla__malla__carrera__coordinacion=9,
                                                                  materia__asignaturamalla__malla=eMalla).exclude(
                            materia_id__in=mat).exclude(materia__asignaturamalla__malla_id__in=eMallasIngles).exclude(
                            id__in=matasig).exclude(materia__asignaturamalla__nivelmalla_id__in=niveles).order_by(
                            'materia__asignaturamalla__nivelmalla_id', 'materia__asignatura__nombre',
                            'materia__paralelo')
                        if not materias:
                            break
                        for materiaasig in materias:
                            nivel = materiaasig.materia.asignaturamalla.nivelmalla_id
                            if totalasignaturas == 4 and nivel not in niveles:
                                if len(MateriaPlanificacionSedeVirtualExamen.objects.filter(status=True,
                                                                                            asignaturamalla__nivelmalla_id=nivel,
                                                                                            asignaturamalla__malla=eMalla,
                                                                                            aulaplanificacion__turnoplanificacion__fechaplanificacion__fecha=fecha.fecha).values(
                                    'asignaturamalla__nivelmalla_id',
                                    'aulaplanificacion__turnoplanificacion__fechaplanificacion__fecha',
                                    'asignaturamalla_id').distinct().order_by(
                                    'aulaplanificacion__turnoplanificacion__fechaplanificacion__fecha',
                                    'asignaturamalla_id')) == 3:
                                    niveles.append(materiaasig.materia.asignaturamalla.nivelmalla_id)
                                else:
                                    totalasignaturas = 1
                            if materiaasig.materia.asignaturamalla.nivelmalla_id in niveles:
                                totalasignaturas = 0
                                continue
                            if not materia == materiaasig.materia:
                                planificado_materias = 0
                                planificado_materias_aula = 0
                            materia = materiaasig.materia

                            materia_id = materia.id
                            totalmateria = len(materia.asignados_a_esta_materia_todos().exclude(
                                matricula__inscripcion__persona__ppl=True).exclude(
                                matricula__inscripcion__persona__perfilinscripcion__tienediscapacidad=True))
                            # cantidadad_planificadas = aula.cantidadad_planificadas()
                            if not asignaturamalla == materiaasig.materia.asignaturamalla:
                                totalasignaturas += 1
                            asignaturamalla = materiaasig.materia.asignaturamalla
                            if cantidadad_planificadas < capacidad:
                                if materia_id not in mat:
                                    c += 1
                                    matasig.append(materiaasig.pk)
                                    cantidadad_planificadas += 1
                                    planificado_materias += 1
                                    planificado_materias_aula += 1
                                    if planificado_materias == totalmateria and materia_id not in mat:
                                        mat.append(materiaasig.materia_id)
                                        save_data(c, materia, aula, planificado_materias_aula,
                                                  detallemodeloevaluativo_id)
                                        continue
                            else:
                                if materia_id not in mat and planificado_materias_aula > 1:
                                    save_data(c, materia, aula, planificado_materias_aula, detallemodeloevaluativo_id)
                                break
        crear_plan_materias_nivelacion(1, 177, fechasplan_nivelacion)
    except Exception as ex:
        print(ex)


# crear_plan_materias(1, 177, fechasplan)
# crear_planificacion_pregrado_sede_examen_milagro()

# crear_plan_materias(1, 177, fechasplan)

# crear_planificacion_nivelacion_sede_examen_milagro()
# crear_planificacion_nivelacion_sede_examen_santo_domingo()


def asignar_sede_test():
    detallemodeloevaluativo_id = 37
    periodo_id = 177
    limitstodomingo = 6667
    eMallasIngles = Malla.objects.filter(pk__in=[353, 22]).values_list('id', flat=True)
    matriculas = MateriaAsignada.objects.filter(status=True, materia__nivel__periodo_id=periodo_id, retiramateria=False,
                                                matricula__retiradomatricula=False,
                                                matricula__inscripcion__carrera__modalidad=3,
                                                materia__asignaturamalla__malla__carrera__coordinacion__lte=5).exclude(
        materia__asignaturamalla__malla_id__in=eMallasIngles).order_by(
        'materia__asignaturamalla__nivelmalla_id', 'materia__asignatura__nombre', 'materia__paralelo').values_list(
        'matricula_id', flat=True).distinct()
    total = len(matriculas)
    print(total)
    for matricula in matriculas:
        if limitstodomingo == 0:
            sede_id = 1
        else:
            sede_id = 10
        if Matricula.objects.filter(Q(status=True, id=matricula), Q(Q(inscripcion__persona__ppl=True) | Q(
                inscripcion__persona__perfilinscripcion__tienediscapacidad=True) | Q(
                inscripcion__persona__pais_id__gt=1))).exists():
            sede_id = 11
        if not MatriculaSedeExamen.objects.filter(status=True, sede_id=sede_id,
                                                  detallemodeloevaluativo_id=detallemodeloevaluativo_id,
                                                  matricula_id=matricula).exists():
            mat = MatriculaSedeExamen(sede_id=sede_id, detallemodeloevaluativo_id=detallemodeloevaluativo_id,
                                      matricula_id=matricula)
            mat.save()
        # matricu = MatriculaSedeExamen.objects.filter(status=True, sede_id=sede_id, detallemodeloevaluativo_id=detallemodeloevaluativo_id, matricula_id=matricula).first()
        total -= 1
        print(total)
        if limitstodomingo > 0 and sede_id == 10:
            limitstodomingo -= 1
    # asignar_sede_nivelacion_test()


# asignar_sede_test()


def asignar_sede_nivelacion_test():
    detallemodeloevaluativo_id = 114
    periodo_id = 177
    limitstodomingo = 5180
    eMallasIngles = Malla.objects.filter(pk__in=[353, 22]).values_list('id', flat=True)
    matriculas = MateriaAsignada.objects.filter(status=True, materia__nivel__periodo_id=periodo_id, retiramateria=False,
                                                matricula__retiradomatricula=False,
                                                materia__asignaturamalla__malla__carrera__coordinacion=9).exclude(
        materia__asignaturamalla__malla_id__in=eMallasIngles).order_by(
        'materia__asignaturamalla__nivelmalla_id', 'materia__asignatura__nombre', 'materia__paralelo').values_list(
        'matricula_id', flat=True).order_by('matricula_id').distinct()
    total = len(matriculas)
    print(total)
    for matricula in matriculas:
        matr = Matricula.objects.filter(Q(status=True, id=matricula))
        if limitstodomingo == 0:
            sede_id = 1
        else:
            sede_id = 10
        if matr.filter(inscripcion__carrera__modalidad__lte=2):
            sede_id = 1
        if matr.filter(Q(Q(inscripcion__persona__ppl=True) | Q(
                inscripcion__persona__perfilinscripcion__tienediscapacidad=True) | Q(
                inscripcion__persona__pais_id__gt=1))).exists():
            sede_id = 11
        if not MatriculaSedeExamen.objects.filter(status=True, sede_id=sede_id,
                                                  detallemodeloevaluativo_id=detallemodeloevaluativo_id,
                                                  matricula_id=matricula).exists():
            mat = MatriculaSedeExamen(sede_id=sede_id, detallemodeloevaluativo_id=detallemodeloevaluativo_id,
                                      matricula_id=matricula)
            mat.save()
        # matricu = MatriculaSedeExamen.objects.filter(status=True, sede_id=sede_id, detallemodeloevaluativo_id=detallemodeloevaluativo_id, matricula_id=matricula).first()
        # print(matricu)
        total -= 1
        print(total)
        if limitstodomingo > 0 and sede_id == 10:
            limitstodomingo -= 1


# asignar_sede_nivelacion_test()


def crear_planificacion_pregrado_sede_examen_santo_domingo_tsachilas():
    fechas = [
        (datetime(2023, 3, 20, 0, 0, 0)).date(),
        (datetime(2023, 3, 21, 0, 0, 0)).date(),
        (datetime(2023, 3, 22, 0, 0, 0)).date(),
        (datetime(2023, 3, 23, 0, 0, 0)).date(),
        (datetime(2023, 3, 24, 0, 0, 0)).date(),
        (datetime(2023, 3, 25, 0, 0, 0)).date(),
        (datetime(2023, 3, 26, 0, 0, 0)).date(),
        (datetime(2023, 3, 27, 0, 0, 0)).date(),
        (datetime(2023, 3, 28, 0, 0, 0)).date()
    ]

    horaslv = [
        [(datetime(2023, 9, 5, 7, 30, 0)).time(), (datetime(2023, 9, 5, 8, 30, 0)).time()],
        [(datetime(2023, 9, 5, 8, 30, 0)).time(), (datetime(2023, 9, 5, 9, 30, 0)).time()],
        [(datetime(2023, 9, 5, 9, 30, 0)).time(), (datetime(2023, 9, 5, 10, 30, 0)).time()],
        [(datetime(2023, 9, 5, 10, 30, 0)).time(), (datetime(2023, 9, 5, 11, 30, 0)).time()],
        [(datetime(2023, 9, 5, 11, 30, 0)).time(), (datetime(2023, 9, 5, 12, 30, 0)).time()],
        [(datetime(2023, 9, 5, 13, 30, 0)).time(), (datetime(2023, 9, 5, 14, 30, 0)).time()],
        [(datetime(2023, 9, 5, 14, 00, 0)).time(), (datetime(2023, 9, 5, 15, 00, 0)).time()],
        [(datetime(2023, 9, 5, 15, 00, 0)).time(), (datetime(2023, 9, 5, 16, 00, 0)).time()],
        [(datetime(2023, 9, 5, 16, 00, 0)).time(), (datetime(2023, 9, 5, 17, 00, 0)).time()],
        [(datetime(2023, 9, 5, 17, 00, 0)).time(), (datetime(2023, 9, 5, 18, 00, 0)).time()],
        [(datetime(2023, 9, 5, 18, 00, 0)).time(), (datetime(2023, 9, 5, 19, 00, 0)).time()],
        [(datetime(2023, 9, 5, 19, 00, 0)).time(), (datetime(2023, 9, 5, 20, 00, 0)).time()],
        [(datetime(2023, 9, 5, 20, 00, 0)).time(), (datetime(2023, 9, 5, 21, 00, 0)).time()],
        [(datetime(2023, 9, 5, 21, 00, 0)).time(), (datetime(2023, 9, 5, 22, 00, 0)).time()]
    ]
    horasfds = [
        [(datetime(2023, 9, 5, 7, 30, 0)).time(), (datetime(2023, 9, 5, 8, 30, 0)).time()],
        [(datetime(2023, 9, 5, 8, 30, 0)).time(), (datetime(2023, 9, 5, 9, 30, 0)).time()],
        [(datetime(2023, 9, 5, 9, 30, 0)).time(), (datetime(2023, 9, 5, 10, 30, 0)).time()],
        [(datetime(2023, 9, 5, 10, 30, 0)).time(), (datetime(2023, 9, 5, 11, 30, 0)).time()],
        [(datetime(2023, 9, 5, 11, 30, 0)).time(), (datetime(2023, 9, 5, 12, 30, 0)).time()],
        [(datetime(2023, 9, 5, 13, 30, 0)).time(), (datetime(2023, 9, 5, 14, 30, 0)).time()],
        [(datetime(2023, 9, 5, 14, 00, 0)).time(), (datetime(2023, 9, 5, 15, 00, 0)).time()],
        [(datetime(2023, 9, 5, 15, 00, 0)).time(), (datetime(2023, 9, 5, 16, 00, 0)).time()],
        [(datetime(2023, 9, 5, 16, 00, 0)).time(), (datetime(2023, 9, 5, 17, 00, 0)).time()],
        [(datetime(2023, 9, 5, 17, 00, 0)).time(), (datetime(2023, 9, 5, 18, 00, 0)).time()],
        [(datetime(2023, 9, 5, 18, 00, 0)).time(), (datetime(2023, 9, 5, 19, 00, 0)).time()],
        [(datetime(2023, 9, 5, 19, 00, 0)).time(), (datetime(2023, 9, 5, 20, 00, 0)).time()],
        [(datetime(2023, 9, 5, 20, 00, 0)).time(), (datetime(2023, 9, 5, 21, 00, 0)).time()],
        [(datetime(2023, 9, 5, 21, 00, 0)).time(), (datetime(2023, 9, 5, 22, 00, 0)).time()]

    ]
    sede_id = 10
    FechaPlanificacionSedeVirtualExamen.objects.filter(sede_id=sede_id,
                                                       periodo_id=153).delete()
    for fecha in fechas:
        print(f"*** FECHA: {fecha}")
        eFechaPlanificacionSedeVirtualExamenes = FechaPlanificacionSedeVirtualExamen.objects.filter(sede_id=sede_id,
                                                                                                    periodo_id=153,
                                                                                                    fecha=fecha)
        if not eFechaPlanificacionSedeVirtualExamenes.values("id").exists():
            eFechaPlanificacionSedeVirtualExamen = FechaPlanificacionSedeVirtualExamen(sede_id=sede_id, periodo_id=153,
                                                                                       fecha=fecha)
            eFechaPlanificacionSedeVirtualExamen.save()
        eFechaPlanificacionSedeVirtualExamen = eFechaPlanificacionSedeVirtualExamenes.first()

        if eFechaPlanificacionSedeVirtualExamen.fecha.weekday() >= 5:
            horas = horasfds
        else:
            horas = horaslv
        for hora in horas:
            horainicio = hora[0]
            horafin = hora[1]
            print(f"*** FECHA: {fecha} -> hora: {horainicio} - {horafin}")
            eTurnoPlanificacionSedeVirtualExamenes = TurnoPlanificacionSedeVirtualExamen.objects.filter(
                fechaplanificacion=eFechaPlanificacionSedeVirtualExamen, horainicio=horainicio, horafin=horafin)
            if not eTurnoPlanificacionSedeVirtualExamenes.values("id").exists():
                eTurnoPlanificacionSedeVirtualExamen = TurnoPlanificacionSedeVirtualExamen(
                    fechaplanificacion=eFechaPlanificacionSedeVirtualExamen, horainicio=horainicio, horafin=horafin)
                eTurnoPlanificacionSedeVirtualExamen.save()
            else:
                eTurnoPlanificacionSedeVirtualExamen = eTurnoPlanificacionSedeVirtualExamenes.first()

            for eLaboratorioVirtual in LaboratorioVirtual.objects.filter(sedevirtual_id=sede_id,
                                                                         id__in=[280, 279, 131, 130, 129]):
                print(f"*** FECHA: {fecha} -> hora: {horainicio} - {horafin} -> aula: {eLaboratorioVirtual.nombre}")
                eAulaPlanificacionSedeVirtualExamenes = AulaPlanificacionSedeVirtualExamen.objects.filter(
                    turnoplanificacion=eTurnoPlanificacionSedeVirtualExamen, aula=eLaboratorioVirtual)
                if not eAulaPlanificacionSedeVirtualExamenes.values("id").exists():
                    eAulaPlanificacionSedeVirtualExamen = AulaPlanificacionSedeVirtualExamen(
                        turnoplanificacion=eTurnoPlanificacionSedeVirtualExamen, aula=eLaboratorioVirtual)
                    eAulaPlanificacionSedeVirtualExamen.save()
                else:
                    eAulaPlanificacionSedeVirtualExamen = eAulaPlanificacionSedeVirtualExamenes.first()
    print(f"********************************FINALIZA PROCESO PLANIFICAR SEDE SANTO DOMINGO")
    print("")
    print("")
    print("")
    print("")
    print("")
    crear_planificacion()


def planificar_pregrado_unemi_v2_virtual(limite_x_día, sede):
    sede_id = sede
    periodo_id = 153
    detallemodeloevaluativo_id = 37
    materias_completas = []
    cursor = connections['sga_select'].cursor()
    eMallasIngles = Malla.objects.filter(pk__in=[353, 22]).values_list('id', flat=True)
    eMatriculaSedeExamenes = MatriculaSedeExamen.objects.filter(sede_id=sede_id,
                                                                detallemodeloevaluativo_id=detallemodeloevaluativo_id,
                                                                status=True, matricula__status=True,
                                                                matricula__retiradomatricula=False,
                                                                matricula__nivel__periodo_id=periodo_id).distinct()
    # if DEBUG:
    # MateriaAsignadaPlanificacionSedeVirtualExamen.objects.filter(status=True,
    #                                                              aulaplanificacion__turnoplanificacion__fechaplanificacion__periodo_id=periodo_id,
    #                                                              aulaplanificacion__turnoplanificacion__fechaplanificacion__sede_id=sede_id,
    #                                                              detallemodeloevaluativo_id=detallemodeloevaluativo_id).delete()

    eMatriculas = Matricula.objects.filter(pk__in=eMatriculaSedeExamenes.values_list("matricula__id", flat=True),
                                           status=True, retiradomatricula=False, bloqueomatricula=False,
                                           nivel__periodo_id=periodo_id)
    eMatriculas_exclude_ingles = eMatriculas.annotate(
        total_ingles=Count('materiaasignada__materia__asignaturamalla__id', filter=Q(
            materiaasignada__materia__asignaturamalla__malla_id__in=eMallasIngles.values_list('id', flat=True),
            nivel__periodo_id=periodo_id, status=True)),
        total_general=Count('materiaasignada__materia__asignaturamalla__id',
                            filter=Q(nivel__periodo_id=periodo_id, status=True))).filter(
        total_general=F('total_ingles'))
    ids_exclude = []
    ids_exclude.extend(list(eMatriculas_exclude_ingles.values_list('id', flat=True)))
    eMatriculas = eMatriculas.exclude(pk__in=ids_exclude)
    eMallas = Malla.objects.filter(
        pk__in=eMatriculas.values_list('inscripcion__inscripcionmalla__malla_id', flat=True).distinct()).exclude(
        id__in=eMallasIngles)
    # eCarreras = Carrera.objects.filter(pk__in=eMallas.values_list('carrera__id', flat=True))
    # eCoordinaciones = Coordinacion.objects.filter(pk__in=eCarreras.values_list('coordinacion__id', flat=True)).distinct()
    # for eCoordinacion in eCoordinaciones:
    #     eMallas = eMallas.filter(carrera_id__in=eCoordinacion.carreras().values_list('id', flat=True)).order_by('carrera__nombre', 'inicio')
    eMallas = eMallas.order_by('carrera__nombre', 'inicio')
    for eMalla in eMallas:
        eFechaPlanificacionSedeVirtualExamenes = FechaPlanificacionSedeVirtualExamen.objects.filter(sede_id=sede_id,
                                                                                                    periodo_id=periodo_id).order_by(
            'fecha')
        for eFechaPlanificacionSedeVirtualExamen in eFechaPlanificacionSedeVirtualExamenes:
            fecha = eFechaPlanificacionSedeVirtualExamen.fecha
            eTurnoPlanificacionSedeVirtualExamenes = TurnoPlanificacionSedeVirtualExamen.objects.filter(
                fechaplanificacion=eFechaPlanificacionSedeVirtualExamen).order_by('horainicio')
            for eTurnoPlanificacionSedeVirtualExamen in eTurnoPlanificacionSedeVirtualExamenes:
                horainicio = eTurnoPlanificacionSedeVirtualExamen.horainicio
                horafin = eTurnoPlanificacionSedeVirtualExamen.horafin

                # eAulaPlanificacionSedeVirtualExamenes = AulaPlanificacionSedeVirtualExamen.objects.filter(turnoplanificacion__fechaplanificacion__sede_id=sede_id,
                #                                                                                           turnoplanificacion__fechaplanificacion__periodo_id=periodo_id)
                eAulaPlanificacionSedeVirtualExamenes = AulaPlanificacionSedeVirtualExamen.objects.filter(
                    turnoplanificacion=eTurnoPlanificacionSedeVirtualExamen)
                eAulaPlanificacionSedeVirtualExamenes_exclude_llenos = eAulaPlanificacionSedeVirtualExamenes.annotate(
                    total_general=Count('materiaasignadaplanificacionsedevirtualexamen__id', filter=Q(
                        materiaasignadaplanificacionsedevirtualexamen__materiaasignada__matricula__nivel__periodo_id=periodo_id,
                        status=True))).filter(total_general=F('aula__capacidad'))
                eAulaPlanificacionSedeVirtualExamenes.exclude(
                    pk__in=eAulaPlanificacionSedeVirtualExamenes_exclude_llenos.values_list("id", flat=True))
                eAulaPlanificacionSedeVirtualExamenes = eAulaPlanificacionSedeVirtualExamenes.order_by(
                    'turnoplanificacion__fechaplanificacion__fecha',
                    'turnoplanificacion__horainicio').distinct()
                # totalAulaSinLlenar = len(eAulaPlanificacionSedeVirtualExamenes.values("id"))
                contadorAulaSinLlenar = 0
                # banderaBreakAula = False
                for eAulaPlanificacionSedeVirtualExamen in eAulaPlanificacionSedeVirtualExamenes:
                    eAula = eAulaPlanificacionSedeVirtualExamen.aula
                    capacidad = eAula.capacidad
                    cantidadad_planificadas = eAulaPlanificacionSedeVirtualExamen.cantidadad_planificadas()
                    eTurnoPlanificacionSedeVirtualExamen = eAulaPlanificacionSedeVirtualExamen.turnoplanificacion
                    # horainicio = eTurnoPlanificacionSedeVirtualExamen.horainicio
                    # horafin = eTurnoPlanificacionSedeVirtualExamen.horafin
                    eFechaPlanificacionSedeVirtualExamen = eTurnoPlanificacionSedeVirtualExamen.fechaplanificacion
                    # fecha = eFechaPlanificacionSedeVirtualExamen.fecha
                    if cantidadad_planificadas < capacidad:
                        print(
                            f"Se procede a planificar en la fecha {fecha} en el horario {horainicio} a {horafin} en el aula {eAula.nombre}")
                        eMateriaAsignadaPlanificacionSedeVirtualExamenes = MateriaAsignadaPlanificacionSedeVirtualExamen.objects.filter(
                            status=True,
                            aulaplanificacion__turnoplanificacion__fechaplanificacion__periodo_id=periodo_id,
                            detallemodeloevaluativo_id=detallemodeloevaluativo_id).values_list("materiaasignada__id",
                                                                                               flat=True)
                        eMateriaAsignadaPlanificacionSedeVirtualExamenesmatriculas = MateriaAsignadaPlanificacionSedeVirtualExamen.objects.filter(
                            status=True,
                            aulaplanificacion__turnoplanificacion__fechaplanificacion__periodo_id=periodo_id,
                            detallemodeloevaluativo_id=detallemodeloevaluativo_id).values_list(
                            "materiaasignada__matricula__id", flat=True)

                        filter_conflicto = (Q(aulaplanificacion__turnoplanificacion__horainicio__lte=horafin,
                                              aulaplanificacion__turnoplanificacion__horafin__gte=horafin,
                                              aulaplanificacion__turnoplanificacion__fechaplanificacion__fecha=fecha) |
                                            Q(aulaplanificacion__turnoplanificacion__horainicio__lte=horainicio,
                                              aulaplanificacion__turnoplanificacion__horafin__gte=horainicio,
                                              aulaplanificacion__turnoplanificacion__fechaplanificacion__fecha=fecha))
                        eMatriculas = Matricula.objects.filter(
                            pk__in=eMatriculaSedeExamenes.values_list("matricula__id", flat=True),
                            status=True,
                            retiradomatricula=False,
                            bloqueomatricula=False,
                            nivel__periodo_id=periodo_id,
                            inscripcion__inscripcionmalla__malla=eMalla).exclude(id__in=materias_completas)
                        # eMatriculas = eMatriculas.filter(pk__in=ids_matricula)
                        eMatriculas_exclude_planificadas = eMatriculas.annotate(
                            total_planificadas=Count('materiaasignada__id', filter=Q(
                                materiaasignada__id__in=eMateriaAsignadaPlanificacionSedeVirtualExamenes, status=True),
                                                     nivel__periodo_id=periodo_id, status=True),
                            total_general=Count('materiaasignada__id',
                                                filter=Q(nivel__periodo_id=periodo_id, status=True),
                                                exclude=Q(materiaasignada__materia__asignatura__id=4837) | Q(
                                                    materiaasignada__materia__asignaturamalla__malla_id__in=eMallasIngles.values_list(
                                                        "id", flat=True)))).filter(
                            Q(total_general=F('total_planificadas')))
                        eMatriculas_exclude_planificadas_x_dia = eMatriculas.annotate(
                            total_planificadas=Count('materiaasignada__id', filter=Q(
                                materiaasignada__id__in=eMateriaAsignadaPlanificacionSedeVirtualExamenes.filter(
                                    aulaplanificacion__turnoplanificacion__fechaplanificacion__fecha=fecha),
                                status=True), nivel__periodo_id=periodo_id, status=True)).filter(
                            Q(total_planificadas=limite_x_día))
                        eMatriculas_exclude_ingles = eMatriculas.annotate(
                            total_ingles=Count('materiaasignada__materia__asignaturamalla__id', filter=Q(
                                materiaasignada__materia__asignaturamalla__malla_id__in=eMallasIngles.values_list("id",
                                                                                                                  flat=True),
                                nivel__periodo_id=periodo_id, status=True)),
                            total_general=Count('materiaasignada__materia__asignaturamalla__id',
                                                filter=Q(nivel__periodo_id=periodo_id, status=True))).filter(
                            total_general=F('total_ingles'))
                        ids_exclude = list(
                            eMateriaAsignadaPlanificacionSedeVirtualExamenesmatriculas.filter(filter_conflicto))
                        ids_exclude.extend(list(eMatriculas_exclude_planificadas.values_list('id', flat=True)))
                        ids_exclude.extend(list(eMatriculas_exclude_planificadas_x_dia.values_list('id', flat=True)))
                        ids_exclude.extend(list(eMatriculas_exclude_ingles.values_list('id', flat=True)))
                        eMatriculas = eMatriculas.exclude(pk__in=ids_exclude)
                        sql = f"""SELECT
                                    "sga_matricula"."id",
                                    COUNT("sga_materia"."asignaturamalla_id")
                                            FILTER (WHERE ("sga_asignaturamalla"."malla_id" NOT IN (	SELECT U0."id"
                                                                                                                    FROM "sga_malla" U0
                                                                                                                    WHERE U0."id" IN (353, 22)
                                                                                                                ) AND
                                                                "sga_nivel"."periodo_id" = {periodo_id} AND
                                                                "sga_matricula"."status"
                                                                )
                                                    ) AS "total_general",
                                    COUNT("sga_materia"."asignaturamalla_id")
                                            FILTER (WHERE ("sga_materiaasignada"."id" IN (	SELECT U0."materiaasignada_id"
                                                                                                                    FROM "inno_materiaasignadaplanificacionsedevirtualexamen" U0
                                                                                                                    INNER JOIN "sga_materiaasignada" U1 ON U1."id" = U0.materiaasignada_id
                                                                                                                    WHERE U1."matricula_id" = "sga_matricula"."id"
                                                                                                                ) AND
                                                                "sga_nivel"."periodo_id" = {periodo_id} AND
                                                                "sga_matricula"."status")
                                                    ) AS "total_planificadas"
                                FROM "sga_matricula"
                                INNER JOIN "sga_inscripcion" ON "sga_matricula"."inscripcion_id" = "sga_inscripcion"."id"
                                INNER JOIN "sga_inscripcionmalla" ON "sga_inscripcion"."id" = "sga_inscripcionmalla"."inscripcion_id"
                                INNER JOIN "sga_nivel" ON "sga_matricula"."nivel_id" = "sga_nivel"."id"
                                INNER JOIN "sga_periodo" ON "sga_nivel"."periodo_id" = "sga_periodo"."id"
                                LEFT OUTER JOIN "sga_materiaasignada" ON "sga_matricula"."id" = "sga_materiaasignada"."matricula_id"
                                LEFT OUTER JOIN "sga_materia" ON "sga_materiaasignada"."materia_id" = "sga_materia"."id"
                                LEFT OUTER JOIN "sga_asignaturamalla" ON "sga_materia"."asignaturamalla_id" = "sga_asignaturamalla"."id"
                                WHERE (
                                    NOT "sga_matricula"."bloqueomatricula" AND
                                    "sga_inscripcionmalla"."malla_id" = {eMalla.pk} AND
                                    "sga_nivel"."periodo_id" = {periodo_id} AND
                                    "sga_matricula"."id" IN (
                                                                        SELECT DISTINCT
                                                                            U0."matricula_id"
                                                                        FROM "inno_matriculasedeexamen" U0
                                                                            INNER JOIN "sga_matricula" U2 ON U0."matricula_id" = U2."id"
                                                                            INNER JOIN "sga_nivel" U3 ON U2."nivel_id" = U3."id"
                                                                        WHERE (
                                                                                    U0."detallemodeloevaluativo_id" = {detallemodeloevaluativo_id} AND
                                                                                    U3."periodo_id" = {periodo_id} AND
                                                                                    NOT U2."retiradomatricula" AND
                                                                                    U2."status" AND
                                                                                    U0."sede_id" = {sede_id} AND
                                                                                    U0."status"
                                                                                )
                                                                    ) AND
                                    NOT "sga_matricula"."retiradomatricula" AND
                                    "sga_matricula"."status"
                                    )
                                GROUP BY "sga_matricula"."id"
                                HAVING
                                        COUNT("sga_materia"."asignaturamalla_id")
                                            FILTER (WHERE ("sga_asignaturamalla"."malla_id" NOT IN (	SELECT U0."id"
                                                                                                                    FROM "sga_malla" U0
                                                                                                                    WHERE U0."id" IN (353, 22)
                                                                                                                ) AND
                                                                "sga_nivel"."periodo_id" = {periodo_id} AND
                                                                "sga_matricula"."status"
                                                                )
                                                    )
                                        <>
                                        COUNT("sga_materia"."asignaturamalla_id")
                                            FILTER (WHERE ("sga_materiaasignada"."id" IN (	SELECT U0."materiaasignada_id"
                                                                                                                    FROM "inno_materiaasignadaplanificacionsedevirtualexamen" U0
                                                                                                                    INNER JOIN "sga_materiaasignada" U1 ON U1."id" = U0.materiaasignada_id
                                                                                                                    WHERE U1."matricula_id" = "sga_matricula"."id"
                                                                                                                ) AND
                                                                "sga_nivel"."periodo_id" = {periodo_id} AND
                                                                "sga_matricula"."status")
                                )"""
                        cursor.execute(sql)
                        results = cursor.fetchall()
                        ids_matricula = [r[0] for r in results]
                        eMatriculas = eMatriculas.filter(pk__in=ids_matricula)
                        eMatriculas = eMatriculas.order_by('inscripcion__inscripcionnivel__nivel__orden',
                                                           'inscripcion__persona__apellido1',
                                                           'inscripcion__persona__apellido2',
                                                           'inscripcion__persona__nombres').distinct()
                        contador = cantidadad_planificadas
                        if not eMatriculas.values("id").exists():
                            contadorAulaSinLlenar += 1

                        if contadorAulaSinLlenar > 0:
                            # banderaBreakAula = True
                            break
                        for eMatricula in eMatriculas:
                            if contador >= capacidad:
                                break
                            eMateriaAsignadas = MateriaAsignada.objects.filter(status=True, matricula=eMatricula)
                            eMateriaAsignadas = eMateriaAsignadas.exclude(
                                Q(pk__in=eMateriaAsignadaPlanificacionSedeVirtualExamenes) |
                                Q(materia__asignatura__id=4837) |
                                Q(materia__asignaturamalla__malla_id__in=eMallasIngles.values_list("id", flat=True)))
                            eMateriaAsignadas = eMateriaAsignadas.order_by(
                                'materia__asignaturamalla__nivelmalla__orden')
                            if eMateriaAsignadas.values("id").exists():
                                # eMateriaAsignada = eMateriaAsignadas.first()
                                for eMateriaAsignada in eMateriaAsignadas:
                                    contador += 1
                                    print(
                                        f"------- ({contador}) Se asignada el estudiante {eMateriaAsignada.matricula.inscripcion.persona} en la asignatura {eMateriaAsignada.materia.asignatura.nombre}")
                                    eMateriaAsignadaPlanificacionSedeVirtualExamen = MateriaAsignadaPlanificacionSedeVirtualExamen(
                                        aulaplanificacion=eAulaPlanificacionSedeVirtualExamen,
                                        materiaasignada=eMateriaAsignada,
                                        detallemodeloevaluativo_id=detallemodeloevaluativo_id)
                                    eMateriaAsignadaPlanificacionSedeVirtualExamen.save()
                                    if contador >= capacidad:
                                        break


def crear_planificacion():
    for sede in [11, 10, 1]:
        if sede == 11:
            MateriaAsignadaPlanificacionSedeVirtualExamen.objects.filter(status=True,
                                                                         aulaplanificacion__turnoplanificacion__fechaplanificacion__sede_id=sede,
                                                                         aulaplanificacion__turnoplanificacion__fechaplanificacion__periodo_id=153,
                                                                         detallemodeloevaluativo_id=37).delete()
            planificar_pregrado_unemi_v2_virtual(10, sede)
        else:
            MateriaAsignadaPlanificacionSedeVirtualExamen.objects.filter(status=True,
                                                                         aulaplanificacion__turnoplanificacion__fechaplanificacion__sede_id=sede,
                                                                         aulaplanificacion__turnoplanificacion__fechaplanificacion__periodo_id=153,
                                                                         detallemodeloevaluativo_id=37).delete()
            planificar_pregrado_unemi_v2(4, sede)


fechas_horario = (
    ('2023-01-13', '2023-03-20'),
    ('2023-01-16', '2023-03-20'),
    ('2023-01-17', '2023-03-21'),
    ('2023-01-18', '2023-03-22'),
    ('2023-01-19', '2023-03-23'),
    ('2023-01-20', '2023-03-24'),
    ('2023-01-23', '2023-03-25'),
    ('2023-01-24', '2023-03-26'),
    ('2023-01-25', '2023-03-27'),
    ('2023-01-26', '2023-03-28'),
    ('2023-01-30', '2023-03-25'),
    ('2023-01-31', '2023-03-26'),
    ('2023-02-01', '2023-03-27'),
    ('2023-02-02', '2023-03-28'),
    ('2023-01-27', '2023-03-20'),
    ('2023-02-03', '2023-03-21'),
)


def examanes_copia_2_parcial(fechas_horario):
    with transaction.atomic():
        try:
            for fecha in fechas_horario:
                print(fecha[0])
                for horario in HorarioExamen.objects.filter(status=True, fecha=fecha[0], materia__nivel__periodo_id=153,
                                                            materia__asignaturamalla__malla__carrera__modalidad=3):
                    materia = Materia.objects.get(pk=horario.materia_id)
                    print(materia.asignatura.nombre)
                    if not HorarioExamen.objects.filter(materia=materia,
                                                        detallemodelo_id=37,
                                                        fecha=fecha[1],
                                                        turno=horario.turno).exists():
                        horarioexamen = HorarioExamen(materia=materia,
                                                      detallemodelo_id=37,
                                                      fecha=fecha[1],
                                                      turno=horario.turno)
                        horarioexamen.save()

                        detalle = horario.horarioexamendetalle_set.first()
                        if detalle:
                            if not HorarioExamenDetalle.objects.filter(horarioexamen=horarioexamen).exists():
                                detalleexamen_new = HorarioExamenDetalle(horarioexamen=horarioexamen,
                                                                         horainicio=detalle.horainicio,
                                                                         horafin=detalle.horafin,
                                                                         cantalumnos=detalle.cantalumnos,
                                                                         tiporesponsable=detalle.tiporesponsable,
                                                                         aula=detalle.aula,
                                                                         profesormateria=detalle.profesormateria)

                                detalleexamen_new.save()
                                eAlumnos = MateriaAsignada.objects.filter(status=True, retiramateria=False,
                                                                          materia=materia,
                                                                          matricula__retiradomatricula=False,
                                                                          matricula__status=True).order_by(
                                    'matricula__inscripcion__persona__apellido1',
                                    'matricula__inscripcion__persona__apellido2',
                                    'matricula__inscripcion__persona__nombres').values_list('id', flat=True)
                                eAsignados = HorarioExamenDetalleAlumno.objects.filter(
                                    horarioexamendetalle=detalleexamen_new, status=True,
                                    materiaasignada_id__in=eAlumnos)
                                if eAsignados.values("id").exists():
                                    eAsignados.delete()
                                for eAlumno in eAlumnos:
                                    eHorarioExamenDetalleAlumno = HorarioExamenDetalleAlumno(materiaasignada_id=eAlumno,
                                                                                             horarioexamendetalle=detalleexamen_new)
                                    eHorarioExamenDetalleAlumno.save()
        except Exception as ex:
            transaction.set_rollback(True)
            print(ex)


# examanes_copia_2_parcial(fechas_horario)


# periodo = Periodo.objects.get(id=177)
#
# fecha1 = datetime(2023, 4, 24).date()
# fecha2 = datetime(2022, 4, 30).date()


def segumientos():
    from sga.models import Notificacion, Periodo, Profesor, DetalleDistributivo, PonderacionSeguimiento, Materia, DiapositivaSilaboSemanal, LlamadasMatriculaSeguimientoTutor
    from datetime import datetime
    fecha1 = datetime(2024, 8, 1).date()
    fecha2 = datetime(2024, 8, 31).date()
    periodo = Periodo.objects.get(id=317)
    docentes = DetalleDistributivo.objects.filter(status=True, distributivo__periodo=periodo,
                                                  criteriodocenciaperiodo__criterio_id=136).values_list(
        'distributivo__profesor', flat=True)
    # Materia.objects.filter(profesormateria__profesor=profesor, nivel__periodo=periodo,nivel__periodo__visible=True, profesormateria__status=True,status=True,profesormateria__activo=True).distinct().order_by('asignatura').distinct()
    try:
        c = 1
        print('TOTAL DE DOCENTES  %s' % len(docentes))
        for p in docentes[0:10]:
            profesor = Profesor.objects.get(id=p)

            text = 'N° {}  DOCENTE .... ---  .....: {}'.format(c, profesor)
            print(text)
            for materia in Materia.objects.filter(profesormateria__profesor=profesor, nivel__periodo=periodo,
                                                  nivel__periodo__visible=True, profesormateria__status=True,
                                                  status=True, profesormateria__activo=True,
                                                  inicio__lte=fecha1 + timedelta(days=1)).distinct().order_by(
                'asignatura').distinct():
                print(materia)
                if PonderacionSeguimiento.objects.filter(status=True, tiposeguimiento=1).exists():
                    ponderacion_plataforma = PonderacionSeguimiento.objects.filter(status=True,
                                                                                   tiposeguimiento=1).order_by(
                        '-id').first().porcentaje
                if PonderacionSeguimiento.objects.filter(status=True, tiposeguimiento=2).exists():
                    ponderacion_recurso = PonderacionSeguimiento.objects.filter(status=True,
                                                                                tiposeguimiento=2).order_by(
                        '-id').first().porcentaje
                if PonderacionSeguimiento.objects.filter(status=True, tiposeguimiento=3).exists():
                    ponderacion_actividad = PonderacionSeguimiento.objects.filter(status=True,
                                                                                  tiposeguimiento=3).order_by(
                        '-id').first().porcentaje
                # materia = Materia.objects.get(pk=int(encrypt(request.GET['id'])))
                materiasasignadas = materia.materiaasignada_set.filter(
                    matricula__estado_matricula__in=[2, 3]).order_by(
                    'matricula__inscripcion__persona__apellido1', 'matricula__inscripcion__persona__apellido2')
                inicio = fecha1
                fin = fecha2
                seguimiento = materia.seguimientotutor_set.filter(status=True).order_by('-fechainicio')
                if not seguimiento.filter(fechainicio__range=[inicio, fin]):
                    # hoy = seleccionar_fecha_ramdom()
                    hoy = inicio
                    if seguimiento:
                        seguimientoaux = materia.seguimientotutor_set.filter(status=True, fechainicio__lte=hoy,
                                                                             fechafin__gte=hoy).order_by(
                            '-fechainicio')
                        if seguimientoaux:
                            finic = seguimientoaux[0].fechainicio
                            ffinc = seguimientoaux[0].fechafin
                        else:
                            finic = seguimiento[0].fechafin + timedelta(days=1)
                            ffinc = hoy
                    else:
                        finic = materia.inicio
                        ffinc = hoy

                    lista = []
                    listaalumnos = []
                    if materia.tareas_asignatura_moodle(finic, ffinc):
                        listaidtareas = []
                        for listatarea in materia.tareas_asignatura_moodle(finic, ffinc):
                            listaidtareas.append(listatarea[0])
                        lista.append(listaidtareas)
                    else:
                        lista.append(0)
                    if materia.foros_asignatura_moodledos(finic, ffinc):
                        listaidforum = []
                        for listaforo in materia.foros_asignatura_moodledos(finic, ffinc):
                            listaidforum.append(listaforo[0])
                        lista.append(listaidforum)
                    else:
                        lista.append(0)
                    if materia.test_asignatura_moodle(finic, ffinc):
                        listaidtest = []
                        for listatest in materia.test_asignatura_moodle(finic, ffinc):
                            listaidtest.append(listatest)
                        lista.append(listaidtest)
                    else:
                        lista.append(0)
                    diapositivas = DiapositivaSilaboSemanal.objects.filter(silabosemanal__silabo__materia=materia,
                                                                           silabosemanal__fechafinciosemana__range=(
                                                                               finic, ffinc),
                                                                           silabosemanal__silabo__status=True,
                                                                           iddiapositivamoodle__gt=0)
                    if diapositivas:
                        listaidpresentacion = []
                        for listadias in diapositivas:
                            listaidpresentacion.append(listadias.iddiapositivamoodle)
                        lista.append(listaidpresentacion)
                    else:
                        lista.append(0)
                    fechacalcula = datetime.strptime('2020-07-24', '%Y-%m-%d').date()
                    if finic > fechacalcula:
                        guiasestudiantes = GuiaEstudianteSilaboSemanal.objects.filter(
                            silabosemanal__silabo__materia=materia,
                            silabosemanal__fechafinciosemana__range=(
                                finic, ffinc), silabosemanal__silabo__status=True,
                            idguiaestudiantemoodle__gt=0)
                        if guiasestudiantes:
                            listaidguiaestudiante = []
                            for listaguiasestu in guiasestudiantes:
                                listaidguiaestudiante.append(listaguiasestu.idguiaestudiantemoodle)
                            lista.append(listaidguiaestudiante)
                        else:
                            lista.append(0)
                        guiasdocentes = GuiaDocenteSilaboSemanal.objects.filter(
                            silabosemanal__silabo__materia=materia,
                            silabosemanal__fechafinciosemana__range=(finic, ffinc),
                            silabosemanal__silabo__status=True,
                            idguiadocentemoodle__gt=0)
                        if guiasdocentes:
                            listaidguiadocente = []
                            for listaguiasdoce in guiasdocentes:
                                listaidguiadocente.append(listaguiasdoce.idguiadocentemoodle)
                            lista.append(listaidguiadocente)
                        else:
                            lista.append(0)
                        compendios = CompendioSilaboSemanal.objects.filter(silabosemanal__silabo__materia=materia,
                                                                           silabosemanal__fechafinciosemana__range=(
                                                                               finic, ffinc),
                                                                           silabosemanal__silabo__status=True,
                                                                           idmcompendiomoodle__gt=0)
                        if compendios:
                            listaidcompendio = []
                            for listacompendios in compendios:
                                listaidcompendio.append(listacompendios.idmcompendiomoodle)
                            lista.append(listaidcompendio)
                        else:
                            lista.append(0)
                    else:
                        lista.append(0)
                        lista.append(0)
                        lista.append(0)
                    totalverde = 0
                    totalamarillo = 0
                    totalrojo = 0
                    for alumnos in materiasasignadas:
                        nombres = alumnos.matricula.inscripcion.persona.apellido1 + ' ' + alumnos.matricula.inscripcion.persona.apellido2 + ' ' + alumnos.matricula.inscripcion.persona.nombres
                        esppl = 'NO'
                        esdiscapacidad = 'NO'
                        if alumnos.matricula.inscripcion.persona.ppl:
                            esppl = 'SI'
                        if alumnos.matricula.inscripcion.persona.mi_perfil().tienediscapacidad:
                            esdiscapacidad = 'SI'
                        if alumnos.matricula.nivel.periodo_id >= 113:
                            totalaccesologuin = float("{:.2f}".format(
                                alumnos.matricula.inscripcion.persona.total_loguinusermoodle_sin_findesemana(finic,
                                                                                                             ffinc)))
                        else:
                            totalaccesologuin = float(
                                "{:.2f}".format(
                                    alumnos.matricula.inscripcion.persona.total_loguinusermoodle(finic, ffinc)))
                        totalaccesorecurso = float("{:.2f}".format(
                            alumnos.matricula.inscripcion.persona.total_accesorecursomoodle(finic, ffinc,
                                                                                            materia.idcursomoodle,
                                                                                            lista)))
                        totalcumplimiento = float("{:.2f}".format(
                            alumnos.matricula.inscripcion.persona.total_cumplimientomoodle(finic, ffinc,
                                                                                           materia.idcursomoodle,
                                                                                           lista)))
                        totalporcentaje = float(
                            "{:.2f}".format(((((totalaccesologuin * ponderacion_plataforma) / 100) + (
                                    (totalaccesorecurso * ponderacion_recurso) / 100) + (
                                                      (totalcumplimiento * ponderacion_actividad) / 100)))))

                        if totalporcentaje >= 70:
                            colorfondo = '5bb75b'
                            totalverde += 1
                        if totalporcentaje <= 30:
                            colorfondo = 'b94a48'
                            totalrojo += 1
                        if totalporcentaje > 31 and totalporcentaje < 70:
                            colorfondo = 'faa732'
                            totalamarillo += 1
                        listaalumnos.append([alumnos.matricula.inscripcion.persona.cedula,
                                             nombres,
                                             esppl,
                                             esdiscapacidad,
                                             totalaccesologuin,
                                             totalaccesorecurso,
                                             totalcumplimiento,
                                             totalporcentaje,
                                             alumnos.matricula.inscripcion.persona.email,
                                             alumnos.matricula.inscripcion.persona.telefono,
                                             alumnos.matricula.inscripcion.persona.canton,
                                             colorfondo,
                                             alumnos.matricula.inscripcion.id,
                                             alumnos.matricula,
                                             alumnos.esta_retirado(),
                                             alumnos.id,
                                             alumnos.retiromanual])

                    try:
                        porcentajeverde = float("{:.2f}".format((totalverde / materiasasignadas.count()) * 100))
                    except ZeroDivisionError:
                        porcentajeverde = 0
                    try:
                        porcentajerojo = float("{:.2f}".format((totalrojo / materiasasignadas.count()) * 100))
                    except ZeroDivisionError:
                        porcentajerojo = 0
                    try:
                        porcentajeamarillo = float("{:.2f}".format((totalamarillo / materiasasignadas.count()) * 100))
                    except ZeroDivisionError:
                        porcentajeamarillo = 0

                    s = SeguimientoTutor.objects.filter(materia=materia, tutor=profesor, fechainicio__lte=hoy,
                                                        fechafin__gte=hoy)
                    if not s:
                        s = SeguimientoTutor(tutor=profesor,
                                             fechainicio=finic,
                                             fechafin=ffinc,
                                             materia=materia,
                                             periodo=periodo)
                        s.save()
                    else:
                        s = s[0]
                    # s.matriculaseguimientotutor_set.all().delete()
                    for integrantes in listaalumnos:
                        ppl = False
                        if integrantes[2] == 'SI':
                            ppl = True
                        discapacidad = False
                        if integrantes[3] == 'SI':
                            discapacidad = True
                        if not MatriculaSeguimientoTutor.objects.filter(seguimiento=s,
                                                                        matricula=integrantes[13]).exists():
                            m = MatriculaSeguimientoTutor(seguimiento=s,
                                                          matricula=integrantes[13],
                                                          ppl=ppl,
                                                          discapacidad=discapacidad,
                                                          accesoplataforma=integrantes[4],
                                                          accesorecurso=integrantes[5],
                                                          cumplimientoactividades=integrantes[6],
                                                          promediovariables=integrantes[7],
                                                          color=integrantes[11])
                            m.save()
                        else:
                            m = MatriculaSeguimientoTutor.objects.get(seguimiento=s, matricula=integrantes[13])
                            m.ppl = ppl
                            m.discapacidad = discapacidad
                            m.accesoplataforma = integrantes[4]
                            m.accesorecurso = integrantes[5]
                            m.cumplimientoactividades = integrantes[6]
                            m.promediovariables = integrantes[7]
                            m.color = integrantes[11]
                            m.save()
                        if integrantes[6] < 1000:
                            llamada = LlamadasMatriculaSeguimientoTutor(matriculaseguimientotutor=m,
                                                                        fecha=fecha1,
                                                                        hora=datetime.now().time(),
                                                                        minutos=1,
                                                                        descripcion='Pondrá más atención')
                            llamada.save()
                noti = Notificacion(cuerpo=f'total {profesor}',
                                    titulo='seuiminento',
                                    destinatario_id=33490,
                                    prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                                    tipo=2, en_proceso=False)
                noti.save()
            c += 1


        # asignar_tutorias()

    except Exception as e:
        print(e)
        print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
        noti = Notificacion(cuerpo='Error on line {}'.format(sys.exc_info()[-1].tb_lineno),
                            titulo='error',
                            destinatario_id=33490,
                            prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                            tipo=2, en_proceso=False)
        noti.save()


fecha_materia = (
    ('2023-03-23', 'LENGUAJE', '9:00 A 10:15'),
    ('2023-03-24', 'INVESTIGACI'),
    ('2023-03-25', 'REALIDAD')
)

turno_paralelo = (
    ('A', (datetime(2023, 3, 23, 9, 0, 0)).time(), (datetime(2023, 3, 23, 10, 15, 0)).time()),
    ('B', (datetime(2023, 3, 23, 10, 15, 0)).time(), (datetime(2023, 3, 23, 11, 30, 0)).time()),
    ('B', (datetime(2023, 3, 23, 11, 30, 0)).time(), (datetime(2023, 3, 23, 12, 45, 0)).time()),
    ('C', (datetime(2023, 3, 23, 12, 45, 0)).time(), (datetime(2023, 3, 23, 14, 00, 0)).time()),
    ('C', (datetime(2023, 3, 23, 14, 00, 0)).time(), (datetime(2023, 3, 23, 15, 15, 0)).time()),
    ('S', (datetime(2023, 3, 23, 14, 00, 0)).time(), (datetime(2023, 3, 23, 15, 15, 0)).time()),
    ('S', (datetime(2023, 3, 23, 15, 15, 0)).time(), (datetime(2023, 3, 23, 16, 30, 0)).time()),
)
aulas_id = []


def examanes_transversales(fechas_horario):
    with transaction.atomic():
        try:
            top = 5000
            for fecha in fecha_materia:
                print(fecha[0])
                for materia in Materia.objects.filter(status=True, nivel__periodo_id=153, modeloevaluativo_id=27,
                                                      asignatura__nombre__icontains=fecha_materia[1]).exclude(
                    asignaturamalla__malla__carrera__modalidad=3).order_by('paralelo'):
                    eAlumnos = MateriaAsignada.objects.filter(status=True, retiramateria=False,
                                                              materia=materia,
                                                              matricula__retiradomatricula=False,
                                                              matricula__status=True).order_by(
                        'matricula__inscripcion__persona__apellido1',
                        'matricula__inscripcion__persona__apellido2',
                        'matricula__inscripcion__persona__nombres').values_list('id', flat=True)
                    for turno in turno_paralelo:
                        if turno[0] in materia.paralelo:
                            for aula in Aula.objects.filter(
                                    id__in=[70, 71, 72, 74, 77, 78, 79, 80, 81, 82, 83, 84, 85, 86, 266]):
                                print(materia.asignatura.nombre)
                                if not HorarioExamen.objects.filter(materia=materia,
                                                                    detallemodelo_id=37,
                                                                    fecha=fecha[1],
                                                                    turno=1).exists():
                                    horarioexamen = HorarioExamen(materia=materia,
                                                                  detallemodelo_id=37,
                                                                  fecha=fecha[1],
                                                                  turno=1)
                                    horarioexamen.save()

                                    detalleexamen_new = HorarioExamenDetalle(horarioexamen=horarioexamen,
                                                                             horainicio=turno[1],
                                                                             horafin=turno[2],
                                                                             cantalumnos=len(eAlumnos),
                                                                             aula=aula)

                                    detalleexamen_new.save()

                                    eAsignados = HorarioExamenDetalleAlumno.objects.filter(
                                        horarioexamendetalle=detalleexamen_new, status=True,
                                        materiaasignada_id__in=eAlumnos)
                                    if eAsignados.values("id").exists():
                                        eAsignados.delete()
                                    for eAlumno in eAlumnos:
                                        eHorarioExamenDetalleAlumno = HorarioExamenDetalleAlumno(
                                            materiaasignada_id=eAlumno, horarioexamendetalle=detalleexamen_new)
                                        eHorarioExamenDetalleAlumno.save()

        except Exception as ex:
            transaction.set_rollback(True)
            print(ex)


def set_hora_paralelo(fecha, paralelo, carrera):
    hora = None
    if str(fecha) == '2023-03-23':
        if 'A' in paralelo:
            hora = (datetime(2023, 3, 23, 9, 00, 0)).time()
        elif 'B' in paralelo:
            hora = (datetime(2023, 3, 23, 10, 15, 0)).time()
            if carrera in [138, 140, 141, 158]:
                hora = (datetime(2023, 3, 23, 11, 30, 0)).time()
            elif carrera in [110, 142, 156, 157, 170]:
                hora = (datetime(2023, 3, 23, 12, 45, 0)).time()
        elif 'C' in paralelo:
            hora = (datetime(2023, 3, 23, 14, 00, 0)).time()
            if carrera in [156, 157]:
                hora = (datetime(2023, 3, 23, 11, 30, 0)).time()
        elif 'S' in paralelo:
            hora = (datetime(2023, 3, 23, 15, 15, 0)).time()
    elif str(fecha) == '2023-03-24':
        if 'A' in paralelo:
            hora = (datetime(2023, 3, 23, 9, 00, 0)).time()
        if 'B' in paralelo:
            hora = (datetime(2023, 3, 23, 10, 15, 0)).time()
            if carrera in [142, 170, 90, 129, 157, 1, 110]:
                hora = (datetime(2023, 3, 23, 11, 30, 0)).time()
        if 'C' in paralelo:
            hora = (datetime(2023, 3, 23, 12, 45, 0)).time()
            if carrera in [156, 90, 129, 157]:
                hora = (datetime(2023, 3, 23, 14, 00, 0)).time()
        if 'S' in paralelo:
            hora = (datetime(2023, 3, 23, 15, 00, 0)).time()
            if carrera in [148, 149, 187]:
                hora = (datetime(2023, 3, 23, 15, 15, 0)).time()
    elif str(fecha) == '2023-03-25':
        if 'A' in paralelo:
            hora = (datetime(2023, 3, 23, 9, 00, 0)).time()
        if 'B' in paralelo:
            hora = (datetime(2023, 3, 23, 10, 15, 0)).time()
        if 'C' in paralelo:
            hora = (datetime(2023, 3, 23, 11, 30, 0)).time()
        if 'S' in paralelo:
            hora = (datetime(2023, 3, 23, 11, 30, 0)).time()
            if carrera in [149, 160]:
                hora = (datetime(2023, 3, 23, 11, 30, 0)).time()
    return hora


def planificar_transversales_unemi_v1():
    try:
        materiatxt = 'LENGUAJE'

        print('EJECUNTANDO............')
        sede_id = 1
        periodo_id = 153
        detallemodeloevaluativo_id = 123
        c = 0
        eMatriculaSedeExamenes = MatriculaSedeExamen.objects.filter(sede_id=sede_id,
                                                                    detallemodeloevaluativo_id=detallemodeloevaluativo_id,
                                                                    status=True, matricula__status=True,
                                                                    matricula__retiradomatricula=False,
                                                                    matricula__nivel__periodo_id=periodo_id).distinct().order_by(
            'matricula__inscripcion__carrera__coordinacion', 'matricula__inscripcion__carrera')
        print('TOTAL DE MATRICULADOS', len(eMatriculaSedeExamenes.values('id')))

        eFechaPlanificacionSedeVirtualExamenes = FechaPlanificacionSedeVirtualExamen.objects.filter(sede_id=sede_id,
                                                                                                    periodo_id=periodo_id).order_by(
            'fecha')
        for eFechaPlanificacionSedeVirtualExamen in eFechaPlanificacionSedeVirtualExamenes:
            fecha = eFechaPlanificacionSedeVirtualExamen.fecha
            if str(fecha) == '2023-03-24':
                materiatxt = 'INVESTIGACI'
            if str(fecha) == '2023-03-25':
                materiatxt = 'REALIDAD'
            if str(fecha) == '2023-03-23':
                materiatxt = 'LENGUAJE'
            eTurnoPlanificacionSedeVirtualExamenes = TurnoPlanificacionSedeVirtualExamen.objects.filter(
                fechaplanificacion=eFechaPlanificacionSedeVirtualExamen).order_by('horainicio')
            for coordinacion in eMatriculaSedeExamenes.values_list('matricula__inscripcion__carrera__coordinacion',
                                                                   flat=True).distinct().order_by(
                'matricula__inscripcion__carrera__coordinacion__nombre'):
                for eTurnoPlanificacionSedeVirtualExamen in eTurnoPlanificacionSedeVirtualExamenes.exclude(
                        horainicio__lte='09:00'):
                    horainicio = eTurnoPlanificacionSedeVirtualExamen.horainicio
                    horafin = eTurnoPlanificacionSedeVirtualExamen.horafin
                    eAulaPlanificacionSedeVirtualExamenes = AulaPlanificacionSedeVirtualExamen.objects.filter(
                        turnoplanificacion=eTurnoPlanificacionSedeVirtualExamen)
                    for eAulaPlanificacionSedeVirtualExamen in eAulaPlanificacionSedeVirtualExamenes:
                        eAula = eAulaPlanificacionSedeVirtualExamen.aula
                        capacidad = eAula.capacidad
                        cantidadad_planificadas = eAulaPlanificacionSedeVirtualExamen.cantidadad_planificadas()
                        if cantidadad_planificadas < capacidad:
                            print(
                                f"Se procede a planificar en la fecha {fecha} en el horario {horainicio} a {horafin} en el aula {eAula.nombre}")
                            eMateriaAsignadaPlanificacionSedeVirtualExamenes = MateriaAsignadaPlanificacionSedeVirtualExamen.objects.filter(
                                status=True,
                                aulaplanificacion__turnoplanificacion__fechaplanificacion__periodo_id=periodo_id,
                                detallemodeloevaluativo_id=detallemodeloevaluativo_id)
                            filter_conflicto = (Q(aulaplanificacion__turnoplanificacion__horainicio__lte=horafin,
                                                  aulaplanificacion__turnoplanificacion__horafin__gte=horafin,
                                                  aulaplanificacion__turnoplanificacion__fechaplanificacion__fecha=fecha) |
                                                Q(aulaplanificacion__turnoplanificacion__horainicio__lte=horainicio,
                                                  aulaplanificacion__turnoplanificacion__horafin__gte=horainicio,
                                                  aulaplanificacion__turnoplanificacion__fechaplanificacion__fecha=fecha))
                            eMatriculas = Matricula.objects.filter(
                                pk__in=eMatriculaSedeExamenes.values_list("matricula__id", flat=True),
                                status=True, retiradomatricula=False, termino=True,
                                fechatermino__isnull=False, bloqueomatricula=False,
                                nivel__periodo_id=periodo_id, inscripcion__carrera__coordinacion=coordinacion)
                            # eMatriculas = eMatriculas.exclude(pk__in=eMateriaAsignadaPlanificacionSedeVirtualExamenes.values_list("materiaasignada__matricula__id", flat=True).filter(filter_conflicto))
                            # eMatriculas_exclude_ingles = eMatriculas.annotate(total_ingles=Count('materiaasignada__materia__asignaturamalla__malla_id', filter=Q(materiaasignada__materia__asignaturamalla__malla__in=eMallaIngles, nivel__periodo_id=periodo_id, status=True)),
                            #                                                   total_general=Count('materiaasignada__materia__asignaturamalla__malla_id', filter=Q(nivel__periodo_id=periodo_id, status=True))).filter(total_general=F('total_ingles'))
                            eMatriculas_exclude_planificadas = eMatriculas.annotate(
                                total_planificadas=Count('materiaasignada__id', filter=Q(
                                    materiaasignada__id__in=eMateriaAsignadaPlanificacionSedeVirtualExamenes.values_list(
                                        "materiaasignada__id", flat=True), status=True), nivel__periodo_id=periodo_id,
                                                         status=True),
                                total_general=Count('materiaasignada__id',
                                                    filter=Q(nivel__periodo_id=periodo_id, status=True),
                                                    exclude=Q(materiaasignada__materia__asignatura__id=4837))).filter(
                                total_general=F('total_planificadas'))
                            # eMatriculas = eMatriculas.exclude(pk__in=eMatriculas_exclude_ingles.values_list('id', flat=True))
                            # eMatriculas = eMatriculas.exclude(pk__in=eMatriculas_exclude_planificadas.values_list('id', flat=True))
                            ids_exclude = list(eMateriaAsignadaPlanificacionSedeVirtualExamenes.values_list(
                                "materiaasignada__matricula__id", flat=True).filter(filter_conflicto))
                            # ids_exclude.extend(list(eMatriculas_exclude_ingles.values_list('id', flat=True)))
                            ids_exclude.extend(list(eMatriculas_exclude_planificadas.values_list('id', flat=True)))
                            eMatriculas = eMatriculas.exclude(pk__in=ids_exclude)
                            # eMatriculas = eMatriculas.order_by('inscripcion__persona__apellido1',
                            #                                    'inscripcion__persona__apellido2',
                            #                                    'inscripcion__persona__nombres').distinct()
                            # eMatriculas = eMatriculas.exclude(Q(materiaasignada__id__in=eMateriaAsignadaPlanificacionSedeVirtualExamenes.values_list("materiaasignada__id", flat=True)) | Q(materiaasignada__materia__asignatura__id=4837))
                            contador = cantidadad_planificadas
                            matriculas_materia = MateriaAsignada.objects.filter(status=True,
                                                                                materia__asignatura__nombre__icontains=materiatxt,
                                                                                matricula_id__in=eMatriculas,
                                                                                materia__nivel__periodo_id=153,
                                                                                materia__modeloevaluativo_id=27).exclude(
                                Q(pk__in=eMateriaAsignadaPlanificacionSedeVirtualExamenes.values_list(
                                    "materiaasignada__id", flat=True)) | Q(materia__asignatura__id=4837)).values_list(
                                'matricula_id', flat=True).exclude(materia__paralelo__icontains='A')
                            if matriculas_materia.exists():
                                matriculas_materia = matriculas_materia.order_by('materia__paralelo',
                                                                                 'matricula__inscripcion__carrera__nombre')
                                for eMatricula in matriculas_materia:
                                    eMateriaAsignadas = MateriaAsignada.objects.filter(status=True,
                                                                                       matricula_id=eMatricula,
                                                                                       materia__asignatura__nombre__icontains=materiatxt,
                                                                                       materia__nivel__periodo_id=153,
                                                                                       materia__modeloevaluativo_id=27).exclude(
                                        Q(pk__in=eMateriaAsignadaPlanificacionSedeVirtualExamenes.values_list(
                                            "materiaasignada__id", flat=True)) | Q(
                                            materia__asignatura__id=4837)).exclude(materia__paralelo__icontains='A')
                                    if eMateriaAsignadas.values("id").exists():
                                        eMateriaAsignada = eMateriaAsignadas.first()
                                        if eMateriaAsignada.matricula.inscripcion.persona.cedula == '0921366241':
                                            print(12)
                                        horai = set_hora_paralelo(fecha, eMateriaAsignada.materia.paralelo,
                                                                  Matricula.objects.get(
                                                                      id=eMatricula).inscripcion.carrera_id)
                                        if horai:
                                            if horainicio >= horai:
                                                contador += 1
                                                print(
                                                    f"------- ({contador}) Se asignada el estudiante {eMateriaAsignada.matricula.inscripcion.persona} en la asignatura {eMateriaAsignada.materia.asignatura.nombre}")
                                                eMateriaAsignadaPlanificacionSedeVirtualExamen = MateriaAsignadaPlanificacionSedeVirtualExamen(
                                                    aulaplanificacion=eAulaPlanificacionSedeVirtualExamen,
                                                    materiaasignada=eMateriaAsignada,
                                                    detallemodeloevaluativo_id=detallemodeloevaluativo_id)
                                                eMateriaAsignadaPlanificacionSedeVirtualExamen.save()
                                                if contador >= capacidad:
                                                    break
            c += 1
    except Exception as e:
        print(e)


def crear_planificacion_examanes_transversales_milagro():
    fechas = [
        (datetime(2023, 3, 23, 0, 0, 0)).date(),
        (datetime(2023, 3, 24, 0, 0, 0)).date(),
        (datetime(2023, 3, 25, 0, 0, 0)).date(),
    ]

    horas = [
        [(datetime(2023, 3, 23, 9, 0, 0)).time(), (datetime(2023, 3, 23, 10, 15, 0)).time()],
        [(datetime(2023, 3, 23, 10, 15, 0)).time(), (datetime(2023, 3, 23, 11, 30, 0)).time()],
        [(datetime(2023, 3, 23, 11, 30, 0)).time(), (datetime(2023, 3, 23, 12, 45, 0)).time()],
        [(datetime(2023, 3, 23, 12, 45, 0)).time(), (datetime(2023, 3, 23, 14, 00, 0)).time()],
        [(datetime(2023, 3, 23, 14, 00, 0)).time(), (datetime(2023, 3, 23, 15, 15, 0)).time()],
        [(datetime(2023, 3, 23, 15, 15, 0)).time(), (datetime(2023, 3, 23, 16, 30, 0)).time()],
        [(datetime(2023, 3, 23, 16, 30, 0)).time(), (datetime(2023, 3, 23, 17, 15, 0)).time()]
    ]
    sede_id = 1
    for fecha in fechas:
        print(f"*** FECHA: {fecha}")
        eFechaPlanificacionSedeVirtualExamenes = FechaPlanificacionSedeVirtualExamen.objects.filter(sede_id=sede_id,
                                                                                                    periodo_id=153,
                                                                                                    fecha=fecha)
        if not eFechaPlanificacionSedeVirtualExamenes.values("id").exists():
            eFechaPlanificacionSedeVirtualExamen = FechaPlanificacionSedeVirtualExamen(sede_id=sede_id,
                                                                                       periodo_id=153,
                                                                                       fecha=fecha)
            eFechaPlanificacionSedeVirtualExamen.save()
        else:
            eFechaPlanificacionSedeVirtualExamen = eFechaPlanificacionSedeVirtualExamenes.first()

        for hora in horas:
            horainicio = hora[0]
            horafin = hora[1]
            print(f"*** FECHA: {fecha} -> hora: {horainicio} - {horafin}")
            eTurnoPlanificacionSedeVirtualExamenes = TurnoPlanificacionSedeVirtualExamen.objects.filter(
                fechaplanificacion=eFechaPlanificacionSedeVirtualExamen, horainicio=horainicio, horafin=horafin)
            if not eTurnoPlanificacionSedeVirtualExamenes.values("id").exists():
                eTurnoPlanificacionSedeVirtualExamen = TurnoPlanificacionSedeVirtualExamen(
                    fechaplanificacion=eFechaPlanificacionSedeVirtualExamen,
                    horainicio=horainicio,
                    horafin=horafin)
                eTurnoPlanificacionSedeVirtualExamen.save()
            else:
                eTurnoPlanificacionSedeVirtualExamen = eTurnoPlanificacionSedeVirtualExamenes.first()

            for eLaboratorioVirtual in LaboratorioVirtual.objects.filter(sedevirtual_id=sede_id, activo=True):
                print(f"*** FECHA: {fecha} -> hora: {horainicio} - {horafin} -> aula: {eLaboratorioVirtual.nombre}")
                eAulaPlanificacionSedeVirtualExamenes = AulaPlanificacionSedeVirtualExamen.objects.filter(
                    turnoplanificacion=eTurnoPlanificacionSedeVirtualExamen, aula=eLaboratorioVirtual)
                if not eAulaPlanificacionSedeVirtualExamenes.values("id").exists():
                    eAulaPlanificacionSedeVirtualExamen = AulaPlanificacionSedeVirtualExamen(
                        turnoplanificacion=eTurnoPlanificacionSedeVirtualExamen,
                        aula=eLaboratorioVirtual)
                    eAulaPlanificacionSedeVirtualExamen.save()
                else:
                    eAulaPlanificacionSedeVirtualExamen = eAulaPlanificacionSedeVirtualExamenes.first()
    print(f"********************************FINALIZA PROCESO PLANIFICAR SEDE MILAGRO")
    print("")
    print("")
    print("")
    print("")
    print("")
    planificar_transversales_unemi_v1()


# planificar_transversales_unemi_v1()


def verificar_materias_matricula():
    eMatriculaSedeExamenes = MatriculaSedeExamen.objects.filter(sede_id=1, detallemodeloevaluativo_id=123,
                                                                status=True, matricula__status=True,
                                                                matricula__retiradomatricula=False,
                                                                matricula__nivel__periodo_id=153).distinct().order_by(
        'id')
    for eMatricula in eMatriculaSedeExamenes:
        eMateriaAsignadas = MateriaAsignada.objects.filter(status=True, matricula=eMatricula.matricula,
                                                           materia__nivel__periodo_id=153,
                                                           materia__modeloevaluativo_id=27)
        for materia in eMateriaAsignadas:
            planificado = MateriaAsignadaPlanificacionSedeVirtualExamen.objects.filter(status=True,
                                                                                       materiaasignada=materia,
                                                                                       aulaplanificacion__turnoplanificacion__fechaplanificacion__periodo_id=153,
                                                                                       detallemodeloevaluativo_id=123).exists()

            if not planificado:
                print(materia.materia.asignatura.nombre, 'NO PLANIFICADO')


# verificar_materias_matricula()
grupo_turnos = (
    ((datetime(2023, 3, 23, 9, 0, 0)),
     (datetime(2023, 3, 23, 10, 15, 0)),
     (datetime(2023, 3, 23, 11, 30, 0))),
    ((datetime(2023, 3, 23, 12, 45, 0)),
     (datetime(2023, 3, 23, 14, 00, 0)),
     (datetime(2023, 3, 23, 15, 15, 0)),
     (datetime(2023, 3, 23, 16, 30, 0)))
)


def elimina_primer_nivel(periodo, carreras):
    matriculas = Matricula.objects.filter(status=True, nivel__periodo_id=periodo, inscripcion__carrera_id__in=carreras,
                                          nivelmalla_id=1)
    print('total de matriculas: ' + str(len(matriculas)))
    matriculas.delete()
    print('FIN')


def actualiza_nivel(periodo):
    for inscripcion in Inscripcion.objects.filter(matricula__nivel__periodo_id=periodo,
                                                  matricula__retiradomatricula=False):
        inscripcion.actualizar_nivel()
        inscripcion.save()
        print(inscripcion.id)
    print('FIN')


def nuevo():
    matriculas = Matricula.objects.filter(status=True, nivel__periodo_id=177, inscripcion__persona__cedula='0952536936')

    # recorrer las materias de cada matricula
    for matricula in matriculas:
        if matricula.inscripcion.persona.cedula == '0952536936':
            for materia in matricula.materiaasignada_set.filter(status=True):
                if materia.materia.asignaturamalla.asignaturamallapredecesora_set.filter(status=True):
                    asignaturas = []
                    if materia.materia.asignaturamalla.itinerario == matricula.inscripcion.itinerario:
                        asignaturas = materia.materia.asignaturamalla.asignaturamallapredecesora_set.filter(status=True,
                                                                                                            asignaturamalla__itinerario=materia.materia.asignaturamalla.itinerario)
                    elif materia.materia.asignaturamalla.itinerario == 0:
                        asignaturas = materia.materia.asignaturamalla.asignaturamallapredecesora_set.filter(status=True)
                    for eAsignaturaMallaPredecesora in asignaturas:
                        if not matricula.inscripcion.recordacademico_set.values('id').filter(
                                asignatura_id=eAsignaturaMallaPredecesora.predecesora.asignatura_id,
                                aprobada=True, status=True).exists():
                            if Rubro.objects.filter(status=True, matricula=matricula):
                                print(materia, 'Si tiene rubro')
                            else:
                                print(materia)


def desenrolar_alumnos_moodle_pregrado(periodo):
    try:
        from moodle import moodle
        for materia in Materia.objects.filter(status=True, nivel__periodo_id=periodo, cerrado=False,
                                              inglesepunemi=False):
            print(materia)
            print('***********Actualizando Alumnos***********')
            materia.quitar_estudiantes_curso(moodle, 1)
            print('***********Alumnos actualizados***********')
            print('***********Actualizando Docentes***********')
            materia.crear_actualizar_docente_curso(moodle, 1)
            print('***********Docentes actualizados***********')

    except Exception as ex:
        print(ex)


periodo2 = Periodo.objects.get(id=177)

fecha01 = datetime(2023, 5, 1).date()
fecha02 = datetime(2022, 5, 31).date()


def segumientos_tranversal():
    docentes = DetalleDistributivo.objects.filter(status=True, distributivo__periodo=periodo2,
                                                  criteriodocenciaperiodo__criterio_id=164).values_list(
        'distributivo__profesor', flat=True)
    try:
        c = 1
        print('TOTAL DE DOCENTES  %s' % len(docentes))
        for p in docentes:
            profesor = Profesor.objects.get(id=p)
            print('N° {}  DOCENTE .... ---  .....: {}'.format(c, profesor))
            for materia in Materia.objects.filter(profesormateria__profesor=profesor, nivel__periodo=periodo,
                                                  tipomateria=3,
                                                  nivel__periodo__visible=True, profesormateria__status=True,
                                                  status=True, profesormateria__activo=True,
                                                  inicio__lte=fecha1 + timedelta(days=1)).distinct().order_by(
                'asignatura').distinct():
                if not materia.coordinacion().id == 9:
                    print(materia)
                    if PonderacionSeguimiento.objects.filter(status=True, tiposeguimiento=1).exists():
                        ponderacion_plataforma = PonderacionSeguimiento.objects.filter(status=True,
                                                                                       tiposeguimiento=1).order_by(
                            '-id').first().porcentaje
                    if PonderacionSeguimiento.objects.filter(status=True, tiposeguimiento=2).exists():
                        ponderacion_recurso = PonderacionSeguimiento.objects.filter(status=True,
                                                                                    tiposeguimiento=2).order_by(
                            '-id').first().porcentaje
                    if PonderacionSeguimiento.objects.filter(status=True, tiposeguimiento=3).exists():
                        ponderacion_actividad = PonderacionSeguimiento.objects.filter(status=True,
                                                                                      tiposeguimiento=3).order_by(
                            '-id').first().porcentaje
                    # materia = Materia.objects.get(pk=int(encrypt(request.GET['id'])))
                    materiasasignadas = materia.materiaasignada_set.filter(
                        matricula__estado_matricula__in=[2, 3]).order_by(
                        'matricula__inscripcion__persona__apellido1', 'matricula__inscripcion__persona__apellido2')
                    inicio = fecha01
                    fin = fecha02
                    seguimiento = materia.seguimientotutor_set.filter(status=True).order_by('-fechainicio')
                    if not seguimiento.filter(fechainicio__range=[inicio, fin]):
                        # hoy = seleccionar_fecha_ramdom()
                        hoy = inicio
                        if seguimiento:
                            seguimientoaux = materia.seguimientotutor_set.filter(status=True, fechainicio__lte=hoy,
                                                                                 fechafin__gte=hoy).order_by(
                                '-fechainicio')
                            if seguimientoaux:
                                finic = seguimientoaux[0].fechainicio
                                ffinc = seguimientoaux[0].fechafin
                            else:
                                finic = seguimiento[0].fechafin + timedelta(days=1)
                                ffinc = hoy
                        else:
                            finic = materia.inicio
                            ffinc = hoy

                        lista = []
                        listaalumnos = []
                        if materia.tareas_asignatura_moodle(finic, ffinc):
                            listaidtareas = []
                            for listatarea in materia.tareas_asignatura_moodle(finic, ffinc):
                                listaidtareas.append(listatarea[0])
                            lista.append(listaidtareas)
                        else:
                            lista.append(0)
                        if materia.foros_asignatura_moodledos(finic, ffinc):
                            listaidforum = []
                            for listaforo in materia.foros_asignatura_moodledos(finic, ffinc):
                                listaidforum.append(listaforo[0])
                            lista.append(listaidforum)
                        else:
                            lista.append(0)
                        if materia.test_asignatura_moodle(finic, ffinc):
                            listaidtest = []
                            for listatest in materia.test_asignatura_moodle(finic, ffinc):
                                listaidtest.append(listatest)
                            lista.append(listaidtest)
                        else:
                            lista.append(0)
                        diapositivas = DiapositivaSilaboSemanal.objects.filter(silabosemanal__silabo__materia=materia,
                                                                               silabosemanal__fechafinciosemana__range=(
                                                                                   finic, ffinc),
                                                                               silabosemanal__silabo__status=True,
                                                                               iddiapositivamoodle__gt=0)
                        if diapositivas:
                            listaidpresentacion = []
                            for listadias in diapositivas:
                                listaidpresentacion.append(listadias.iddiapositivamoodle)
                            lista.append(listaidpresentacion)
                        else:
                            lista.append(0)
                        fechacalcula = datetime.strptime('2020-07-24', '%Y-%m-%d').date()
                        if finic > fechacalcula:
                            guiasestudiantes = GuiaEstudianteSilaboSemanal.objects.filter(
                                silabosemanal__silabo__materia=materia,
                                silabosemanal__fechafinciosemana__range=(
                                    finic, ffinc), silabosemanal__silabo__status=True,
                                idguiaestudiantemoodle__gt=0)
                            if guiasestudiantes:
                                listaidguiaestudiante = []
                                for listaguiasestu in guiasestudiantes:
                                    listaidguiaestudiante.append(listaguiasestu.idguiaestudiantemoodle)
                                lista.append(listaidguiaestudiante)
                            else:
                                lista.append(0)
                            guiasdocentes = GuiaDocenteSilaboSemanal.objects.filter(
                                silabosemanal__silabo__materia=materia,
                                silabosemanal__fechafinciosemana__range=(finic, ffinc),
                                silabosemanal__silabo__status=True,
                                idguiadocentemoodle__gt=0)
                            if guiasdocentes:
                                listaidguiadocente = []
                                for listaguiasdoce in guiasdocentes:
                                    listaidguiadocente.append(listaguiasdoce.idguiadocentemoodle)
                                lista.append(listaidguiadocente)
                            else:
                                lista.append(0)
                            compendios = CompendioSilaboSemanal.objects.filter(silabosemanal__silabo__materia=materia,
                                                                               silabosemanal__fechafinciosemana__range=(
                                                                                   finic, ffinc),
                                                                               silabosemanal__silabo__status=True,
                                                                               idmcompendiomoodle__gt=0)
                            if compendios:
                                listaidcompendio = []
                                for listacompendios in compendios:
                                    listaidcompendio.append(listacompendios.idmcompendiomoodle)
                                lista.append(listaidcompendio)
                            else:
                                lista.append(0)
                        else:
                            lista.append(0)
                            lista.append(0)
                            lista.append(0)
                        totalverde = 0
                        totalamarillo = 0
                        totalrojo = 0
                        for alumnos in materiasasignadas:
                            nombres = alumnos.matricula.inscripcion.persona.apellido1 + ' ' + alumnos.matricula.inscripcion.persona.apellido2 + ' ' + alumnos.matricula.inscripcion.persona.nombres
                            esppl = 'NO'
                            esdiscapacidad = 'NO'
                            if alumnos.matricula.inscripcion.persona.ppl:
                                esppl = 'SI'
                            if alumnos.matricula.inscripcion.persona.mi_perfil().tienediscapacidad:
                                esdiscapacidad = 'SI'
                            if alumnos.matricula.nivel.periodo_id >= 113:
                                totalaccesologuin = float("{:.2f}".format(
                                    alumnos.matricula.inscripcion.persona.total_loguinusermoodle_sin_findesemana(finic,
                                                                                                                 ffinc)))
                            else:
                                totalaccesologuin = float(
                                    "{:.2f}".format(
                                        alumnos.matricula.inscripcion.persona.total_loguinusermoodle(finic, ffinc)))
                            totalaccesorecurso = float("{:.2f}".format(
                                alumnos.matricula.inscripcion.persona.total_accesorecursomoodle(finic, ffinc,
                                                                                                materia.idcursomoodle,
                                                                                                lista)))
                            totalcumplimiento = float("{:.2f}".format(
                                alumnos.matricula.inscripcion.persona.total_cumplimientomoodle(finic, ffinc,
                                                                                               materia.idcursomoodle,
                                                                                               lista)))
                            totalporcentaje = float(
                                "{:.2f}".format(((((totalaccesologuin * ponderacion_plataforma) / 100) + (
                                        (totalaccesorecurso * ponderacion_recurso) / 100) + (
                                                          (totalcumplimiento * ponderacion_actividad) / 100)))))

                            if totalporcentaje >= 70:
                                colorfondo = '5bb75b'
                                totalverde += 1
                            if totalporcentaje <= 30:
                                colorfondo = 'b94a48'
                                totalrojo += 1
                            if totalporcentaje > 31 and totalporcentaje < 70:
                                colorfondo = 'faa732'
                                totalamarillo += 1
                            listaalumnos.append([alumnos.matricula.inscripcion.persona.cedula,
                                                 nombres,
                                                 esppl,
                                                 esdiscapacidad,
                                                 totalaccesologuin,
                                                 totalaccesorecurso,
                                                 totalcumplimiento,
                                                 totalporcentaje,
                                                 alumnos.matricula.inscripcion.persona.email,
                                                 alumnos.matricula.inscripcion.persona.telefono,
                                                 alumnos.matricula.inscripcion.persona.canton,
                                                 colorfondo,
                                                 alumnos.matricula.inscripcion.id,
                                                 alumnos.matricula,
                                                 alumnos.esta_retirado(),
                                                 alumnos.id,
                                                 alumnos.retiromanual])

                        try:
                            porcentajeverde = float("{:.2f}".format((totalverde / materiasasignadas.count()) * 100))
                        except ZeroDivisionError:
                            porcentajeverde = 0
                        try:
                            porcentajerojo = float("{:.2f}".format((totalrojo / materiasasignadas.count()) * 100))
                        except ZeroDivisionError:
                            porcentajerojo = 0
                        try:
                            porcentajeamarillo = float(
                                "{:.2f}".format((totalamarillo / materiasasignadas.count()) * 100))
                        except ZeroDivisionError:
                            porcentajeamarillo = 0

                        s = SeguimientoTutor.objects.filter(materia=materia, tutor=profesor, fechainicio__lte=hoy,
                                                            fechafin__gte=hoy)
                        if not s:
                            s = SeguimientoTutor(tutor=profesor,
                                                 fechainicio=finic,
                                                 fechafin=ffinc,
                                                 materia=materia,
                                                 periodo=periodo)
                            s.save()
                        else:
                            s = s[0]
                        # s.matriculaseguimientotutor_set.all().delete()
                        for integrantes in listaalumnos:
                            ppl = False
                            if integrantes[2] == 'SI':
                                ppl = True
                            discapacidad = False
                            if integrantes[3] == 'SI':
                                discapacidad = True
                            if not MatriculaSeguimientoTutor.objects.filter(seguimiento=s,
                                                                            matricula=integrantes[13]).exists():
                                m = MatriculaSeguimientoTutor(seguimiento=s,
                                                              matricula=integrantes[13],
                                                              ppl=ppl,
                                                              discapacidad=discapacidad,
                                                              accesoplataforma=integrantes[4],
                                                              accesorecurso=integrantes[5],
                                                              cumplimientoactividades=integrantes[6],
                                                              promediovariables=integrantes[7],
                                                              color=integrantes[11])
                                m.save()
                            else:
                                m = MatriculaSeguimientoTutor.objects.get(seguimiento=s, matricula=integrantes[13])
                                m.ppl = ppl
                                m.discapacidad = discapacidad
                                m.accesoplataforma = integrantes[4]
                                m.accesorecurso = integrantes[5]
                                m.cumplimientoactividades = integrantes[6]
                                m.promediovariables = integrantes[7]
                                m.color = integrantes[11]
                                m.save()
                            if integrantes[6] < 100:
                                llamada = LlamadasMatriculaSeguimientoTutor(matriculaseguimientotutor=m,
                                                                            fecha=fecha1,
                                                                            hora=datetime.now().time(),
                                                                            minutos=1,
                                                                            descripcion='Pondrá más atención')
                                llamada.save()
            c += 1
    except Exception as e:
        print(e)
        print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))


# segumientos_tranversal()


# def check_examan_plan():
#     for fecha in FechaPlanificacionSedeVirtualExamen.objects.filter(status=True, periodo_id=177):
#         total = fecha.turnoplanificacionsedevirtualexamen_set.all

def save_sede_virtual(detallemodeloevaluativo_id, matricula, sede):
    if not MatriculaSedeExamen.objects.filter(status=True, matricula_id=matricula):
        mat = MatriculaSedeExamen(sede_id=sede, detallemodeloevaluativo_id=detallemodeloevaluativo_id,
                                  matricula_id=matricula)
        mat.save()
    else:
        mat = MatriculaSedeExamen.objects.filter(status=True, matricula_id=matricula).first()
        mat.detallemodeloevaluativo_id = detallemodeloevaluativo_id
        mat.sede_id = sede
        mat.save()
    print("SEDE ASIGNADA A: {}".format(mat))

    pass


def asignar_sede_virtual_2023():
    print("Actualizando ppls")
    set_ppl_statatus()
    matriculas = Matricula.objects.filter(
        Q(status=True, nivel__periodo_id=177, retiradomatricula=False, inscripcion__carrera__coordinacion=9,
          inscripcion__carrera__modalidad__in=[1, 2])).exclude(inscripcion__carrera_id__in=[223, 224])

    if matriculas:
        detallemodeloevaluativo_id = 114
        print("Asignando sede a NIVELACION")
        for matricula in matriculas:
            # if matricula.inscripcion.persona.perfilinscripcion_set.first().tienediscapacidad:
            #     if matricula.inscripcion.persona.perfilinscripcion_set.first().tienediscapacidad and matricula.inscripcion.persona.perfilinscripcion_set.first().verificadiscapacidad:
            #         save_sede_virtual(114, matricula.id)
            #     else:
            #         print(" no esta verificado")
            # else:
            save_sede_virtual(detallemodeloevaluativo_id, matricula.id, 1)
    else:
        print("No hay matriculas en NIVELACION")

    # MigrantePersona
    # matriculas_migrante = Matricula.objects.filter(Q(status=True, nivel__periodo_id=177, retiradomatricula=False, inscripcion__carrera__coordinacion__lte=5, inscripcion__persona__migrantepersona__verificado=True, inscripcion__persona__migrantepersona__fecharetorno__gt=datetime.now()))
    # if matriculas_migrante:
    #     for matricula in matriculas_migrante:
    #         save_sede_virtual(37, matricula.id)
    # asig_nivelacion()


def asig_nivelacion():
    matriculas_vitual = Matricula.objects.filter(
        Q(status=True, nivel__periodo_id=177, retiradomatricula=False, inscripcion__carrera__coordinacion__lte=5),
        Q(Q(inscripcion__persona__ppl=True) | Q(
            inscripcion__persona__perfilinscripcion__tienediscapacidad=True)))
    if matriculas_vitual:
        detallemodeloevaluativo_id = 37
        print("Asignando sede a PREGRADO")
        for matricula in matriculas_vitual:
            if matricula.inscripcion.persona.perfilinscripcion_set.first().tienediscapacidad:
                if matricula.inscripcion.persona.perfilinscripcion_set.first().tienediscapacidad and matricula.inscripcion.persona.perfilinscripcion_set.first().verificadiscapacidad:
                    save_sede_virtual(detallemodeloevaluativo_id, matricula.id, 11)
                else:
                    print(" no esta verificado")
            else:
                save_sede_virtual(detallemodeloevaluativo_id, matricula.id, 11)
    # else:
    #     print("No hay matriculas en NIVELACION")

    matriculas_migrante_niv = Matricula.objects.filter(
        Q(status=True, nivel__periodo_id=177, retiradomatricula=False,
          inscripcion__carrera__coordinacion__in=[1, 2, 3, 4, 5],
          inscripcion__persona__migrantepersona__verificado=True))
    if matriculas_migrante_niv:
        for matricula in matriculas_migrante_niv:
            save_sede_virtual(37, matricula.id, 11)


fechasplan = [
    (datetime(2023, 8, 7, 0, 0, 0)).date(),
    (datetime(2023, 8, 8, 0, 0, 0)).date(),
    (datetime(2023, 8, 9, 0, 0, 0)).date(),
    (datetime(2023, 8, 10, 0, 0, 0)).date(),
    (datetime(2023, 8, 14, 0, 0, 0)).date(),
    (datetime(2023, 8, 15, 0, 0, 0)).date(),
    (datetime(2023, 8, 16, 0, 0, 0)).date(),
    (datetime(2023, 8, 17, 0, 0, 0)).date(),
    (datetime(2023, 8, 21, 0, 0, 0)).date(),
    (datetime(2023, 8, 22, 0, 0, 0)).date(),
    (datetime(2023, 8, 23, 0, 0, 0)).date(),
    (datetime(2023, 8, 24, 0, 0, 0)).date(),
    (datetime(2023, 8, 25, 0, 0, 0)).date(),
    (datetime(2023, 8, 26, 0, 0, 0)).date(),
]

fechasplan_nivelacion = [
    (datetime(2023, 7, 24, 0, 0, 0)).date(),
    (datetime(2023, 7, 25, 0, 0, 0)).date(),
    (datetime(2023, 7, 26, 0, 0, 0)).date(),
    (datetime(2023, 7, 27, 0, 0, 0)).date(),
    (datetime(2023, 7, 28, 0, 0, 0)).date()
]

horas_nivelacion = [
    [(datetime(2022, 9, 5, 8, 0, 0)).time(), (datetime(2022, 9, 5, 9, 14, 0)).time()],
    [(datetime(2022, 9, 5, 9, 15, 0)).time(), (datetime(2022, 9, 5, 10, 29, 0)).time()],
    [(datetime(2022, 9, 5, 10, 30, 0)).time(), (datetime(2022, 9, 5, 11, 44, 0)).time()],
    [(datetime(2022, 9, 5, 11, 45, 0)).time(), (datetime(2022, 9, 5, 12, 59, 0)).time()],
    [(datetime(2022, 9, 5, 13, 00, 0)).time(), (datetime(2022, 9, 5, 14, 14, 0)).time()],
    [(datetime(2022, 9, 5, 14, 15, 0)).time(), (datetime(2022, 9, 5, 15, 29, 0)).time()],
    [(datetime(2022, 9, 5, 15, 30, 0)).time(), (datetime(2022, 9, 5, 16, 44, 0)).time()],
    [(datetime(2022, 9, 5, 16, 45, 0)).time(), (datetime(2022, 9, 5, 17, 59, 0)).time()],
    [(datetime(2022, 9, 5, 18, 00, 0)).time(), (datetime(2022, 9, 5, 19, 15, 0)).time()],
]
horas_nivelacion_sto_domingo = [
    [(datetime(2022, 9, 5, 8, 0, 0)).time(), (datetime(2022, 9, 5, 9, 14, 0)).time()],
    [(datetime(2022, 9, 5, 9, 15, 0)).time(), (datetime(2022, 9, 5, 10, 29, 0)).time()],
    [(datetime(2022, 9, 5, 10, 30, 0)).time(), (datetime(2022, 9, 5, 11, 44, 0)).time()],
    [(datetime(2022, 9, 5, 11, 45, 0)).time(), (datetime(2022, 9, 5, 12, 59, 0)).time()],
    [(datetime(2022, 9, 5, 13, 00, 0)).time(), (datetime(2022, 9, 5, 14, 14, 0)).time()],
    [(datetime(2022, 9, 5, 14, 15, 0)).time(), (datetime(2022, 9, 5, 15, 29, 0)).time()],
    [(datetime(2022, 9, 5, 15, 30, 0)).time(), (datetime(2022, 9, 5, 16, 44, 0)).time()]
]
horas_nivelacion_virtual = [
    [(datetime(2022, 9, 5, 8, 0, 0)).time(), (datetime(2022, 9, 5, 16, 44, 0)).time()],
]


def crear_planificacion_admision_sede_examen(fechas, horas, periodo, sede_id):
    for fecha in fechas:
        print(f"*** FECHA: {fecha}")
        eFechaPlanificacionSedeVirtualExamenes = FechaPlanificacionSedeVirtualExamen.objects.filter(sede_id=sede_id,
                                                                                                    periodo_id=periodo,
                                                                                                    fecha=fecha)
        if not eFechaPlanificacionSedeVirtualExamenes.values("id").exists():
            eFechaPlanificacionSedeVirtualExamen = FechaPlanificacionSedeVirtualExamen(sede_id=sede_id,
                                                                                       periodo_id=periodo,
                                                                                       fecha=fecha)
            eFechaPlanificacionSedeVirtualExamen.save()
        else:
            eFechaPlanificacionSedeVirtualExamen = eFechaPlanificacionSedeVirtualExamenes.first()

        for hora in horas:
            horainicio = hora[0]
            horafin = hora[1]
            print(f"*** FECHA: {fecha} -> hora: {horainicio} - {horafin}")
            eTurnoPlanificacionSedeVirtualExamenes = TurnoPlanificacionSedeVirtualExamen.objects.filter(
                fechaplanificacion=eFechaPlanificacionSedeVirtualExamen, horainicio=horainicio, horafin=horafin)
            if not eTurnoPlanificacionSedeVirtualExamenes.values("id").exists():
                eTurnoPlanificacionSedeVirtualExamen = TurnoPlanificacionSedeVirtualExamen(
                    fechaplanificacion=eFechaPlanificacionSedeVirtualExamen,
                    horainicio=horainicio,
                    horafin=horafin)
                eTurnoPlanificacionSedeVirtualExamen.save()
            else:
                eTurnoPlanificacionSedeVirtualExamen = eTurnoPlanificacionSedeVirtualExamenes.first()

            for eLaboratorioVirtual in LaboratorioVirtual.objects.filter(sedevirtual_id=sede_id, activo=True):
                print(f"*** FECHA: {fecha} -> hora: {horainicio} - {horafin} -> aula: {eLaboratorioVirtual.nombre}")
                eAulaPlanificacionSedeVirtualExamenes = AulaPlanificacionSedeVirtualExamen.objects.filter(
                    turnoplanificacion=eTurnoPlanificacionSedeVirtualExamen, aula=eLaboratorioVirtual)
                if not eAulaPlanificacionSedeVirtualExamenes.values("id").exists():
                    eAulaPlanificacionSedeVirtualExamen = AulaPlanificacionSedeVirtualExamen(
                        turnoplanificacion=eTurnoPlanificacionSedeVirtualExamen,
                        aula=eLaboratorioVirtual)
                    eAulaPlanificacionSedeVirtualExamen.save()
                # else:
                #     eAulaPlanificacionSedeVirtualExamen = eAulaPlanificacionSedeVirtualExamenes.first()
    print(f"********************************FINALIZA PROCESO PLANIFICAR SEDE virtual")


def planificar_pregrado_unemi_v2_medicina(limite_x_día, sede, periodo_id, detallemodeloevaluativo_id, fechas):
    try:
        sede_id = sede
        print('PLANIFICANDO SEDE ' + str(sede_id))
        # periodo_id = 153
        # detallemodeloevaluativo_id = 37
        materias_completas = []
        cursor = connections['sga_select'].cursor()
        eMallasIngles = Malla.objects.filter(pk__in=[353, 22]).values_list('id', flat=True)
        eMallasMedicina = Malla.objects.filter(carrera__id__in=[224]).values_list('id', flat=True)
        eMatriculaSedeExamenes = MatriculaSedeExamen.objects.filter(sede_id=sede_id,
                                                                    detallemodeloevaluativo_id=detallemodeloevaluativo_id,
                                                                    status=True, matricula__status=True,
                                                                    matricula__retiradomatricula=False,
                                                                    matricula__nivel__periodo_id=periodo_id).distinct()
        # if DEBUG:
        # MateriaAsignadaPlanificacionSedeVirtualExamen.objects.filter(status=True,
        #                                                              aulaplanificacion__turnoplanificacion__fechaplanificacion__periodo_id=periodo_id,
        #                                                              aulaplanificacion__turnoplanificacion__fechaplanificacion__sede_id=sede_id,
        #                                                              detallemodeloevaluativo_id=detallemodeloevaluativo_id).delete()

        eMatriculas = Matricula.objects.filter(pk__in=eMatriculaSedeExamenes.values_list("matricula__id", flat=True),
                                               status=True, retiradomatricula=False, bloqueomatricula=False,
                                               nivel__periodo_id=periodo_id)
        eMatriculas_exclude_ingles = eMatriculas.annotate(
            total_ingles=Count('materiaasignada__materia__asignaturamalla__id', filter=Q(
                materiaasignada__materia__asignaturamalla__malla_id__in=eMallasIngles.values_list('id', flat=True),
                nivel__periodo_id=periodo_id, status=True)),
            total_general=Count('materiaasignada__materia__asignaturamalla__id',
                                filter=Q(nivel__periodo_id=periodo_id, status=True))).filter(
            total_general=F('total_ingles'))
        ids_exclude = []
        ids_exclude.extend(list(eMatriculas_exclude_ingles.values_list('id', flat=True)))
        eMatriculas = eMatriculas.exclude(pk__in=ids_exclude)
        eMallas = Malla.objects.filter(id__in=eMallasMedicina,
                                       pk__in=eMatriculas.values_list('inscripcion__inscripcionmalla__malla_id',
                                                                      flat=True).distinct()).exclude(
            id__in=eMallasIngles)
        # eCarreras = Carrera.objects.filter(pk__in=eMallas.values_list('carrera__id', flat=True))
        # eCoordinaciones = Coordinacion.objects.filter(pk__in=eCarreras.values_list('coordinacion__id', flat=True)).distinct()
        # for eCoordinacion in eCoordinaciones:
        #     eMallas = eMallas.filter(carrera_id__in=eCoordinacion.carreras().values_list('id', flat=True)).order_by('carrera__nombre', 'inicio')
        eMallas = eMallas.order_by('carrera__nombre', 'inicio')
        for eMalla in eMallas:
            eFechaPlanificacionSedeVirtualExamenes = FechaPlanificacionSedeVirtualExamen.objects.filter(sede_id=sede_id,
                                                                                                        periodo_id=periodo_id,
                                                                                                        fecha__in=fechas).order_by(
                'fecha')
            for eFechaPlanificacionSedeVirtualExamen in eFechaPlanificacionSedeVirtualExamenes:
                fecha = eFechaPlanificacionSedeVirtualExamen.fecha
                eTurnoPlanificacionSedeVirtualExamenes = TurnoPlanificacionSedeVirtualExamen.objects.filter(
                    fechaplanificacion=eFechaPlanificacionSedeVirtualExamen).order_by('horainicio')
                for eTurnoPlanificacionSedeVirtualExamen in eTurnoPlanificacionSedeVirtualExamenes:
                    horainicio = eTurnoPlanificacionSedeVirtualExamen.horainicio
                    horafin = eTurnoPlanificacionSedeVirtualExamen.horafin

                    # eAulaPlanificacionSedeVirtualExamenes = AulaPlanificacionSedeVirtualExamen.objects.filter(turnoplanificacion__fechaplanificacion__sede_id=sede_id,
                    #                                                                                           turnoplanificacion__fechaplanificacion__periodo_id=periodo_id)
                    eAulaPlanificacionSedeVirtualExamenes = AulaPlanificacionSedeVirtualExamen.objects.filter(
                        turnoplanificacion=eTurnoPlanificacionSedeVirtualExamen)
                    eAulaPlanificacionSedeVirtualExamenes_exclude_llenos = eAulaPlanificacionSedeVirtualExamenes.annotate(
                        total_general=Count('materiaasignadaplanificacionsedevirtualexamen__id', filter=Q(
                            materiaasignadaplanificacionsedevirtualexamen__materiaasignada__matricula__nivel__periodo_id=periodo_id,
                            status=True))).filter(total_general=F('aula__capacidad'))
                    eAulaPlanificacionSedeVirtualExamenes.exclude(
                        pk__in=eAulaPlanificacionSedeVirtualExamenes_exclude_llenos.values_list("id", flat=True))
                    eAulaPlanificacionSedeVirtualExamenes = eAulaPlanificacionSedeVirtualExamenes.order_by(
                        'turnoplanificacion__fechaplanificacion__fecha',
                        'turnoplanificacion__horainicio').distinct()
                    # totalAulaSinLlenar = len(eAulaPlanificacionSedeVirtualExamenes.values("id"))
                    contadorAulaSinLlenar = 0
                    # banderaBreakAula = False
                    for eAulaPlanificacionSedeVirtualExamen in eAulaPlanificacionSedeVirtualExamenes:
                        eAula = eAulaPlanificacionSedeVirtualExamen.aula
                        capacidad = eAula.capacidad
                        cantidadad_planificadas = eAulaPlanificacionSedeVirtualExamen.cantidadad_planificadas()
                        eTurnoPlanificacionSedeVirtualExamen = eAulaPlanificacionSedeVirtualExamen.turnoplanificacion
                        # horainicio = eTurnoPlanificacionSedeVirtualExamen.horainicio
                        # horafin = eTurnoPlanificacionSedeVirtualExamen.horafin
                        eFechaPlanificacionSedeVirtualExamen = eTurnoPlanificacionSedeVirtualExamen.fechaplanificacion
                        # fecha = eFechaPlanificacionSedeVirtualExamen.fecha
                        if cantidadad_planificadas < capacidad:
                            print(
                                f"Se procede a planificar en la fecha {fecha} en el horario {horainicio} a {horafin} en el aula {eAula.nombre}")
                            eMateriaAsignadaPlanificacionSedeVirtualExamenes = MateriaAsignadaPlanificacionSedeVirtualExamen.objects.filter(
                                status=True,
                                aulaplanificacion__turnoplanificacion__fechaplanificacion__periodo_id=periodo_id,
                                detallemodeloevaluativo_id=detallemodeloevaluativo_id).values_list(
                                "materiaasignada__id",
                                flat=True)
                            eMateriaAsignadaPlanificacionSedeVirtualExamenesmatriculas = MateriaAsignadaPlanificacionSedeVirtualExamen.objects.filter(
                                status=True,
                                aulaplanificacion__turnoplanificacion__fechaplanificacion__periodo_id=periodo_id,
                                detallemodeloevaluativo_id=detallemodeloevaluativo_id).values_list(
                                "materiaasignada__matricula__id", flat=True)

                            filter_conflicto = (Q(aulaplanificacion__turnoplanificacion__horainicio__lte=horafin,
                                                  aulaplanificacion__turnoplanificacion__horafin__gte=horafin,
                                                  aulaplanificacion__turnoplanificacion__fechaplanificacion__fecha=fecha) |
                                                Q(aulaplanificacion__turnoplanificacion__horainicio__lte=horainicio,
                                                  aulaplanificacion__turnoplanificacion__horafin__gte=horainicio,
                                                  aulaplanificacion__turnoplanificacion__fechaplanificacion__fecha=fecha))
                            eMatriculas = Matricula.objects.filter(
                                pk__in=eMatriculaSedeExamenes.values_list("matricula__id", flat=True),
                                status=True,
                                retiradomatricula=False,
                                bloqueomatricula=False,
                                nivel__periodo_id=periodo_id,
                                inscripcion__inscripcionmalla__malla=eMalla).exclude(id__in=materias_completas)
                            # eMatriculas = eMatriculas.filter(pk__in=ids_matricula)
                            eMatriculas_exclude_planificadas = eMatriculas.annotate(
                                total_planificadas=Count('materiaasignada__id', filter=Q(
                                    materiaasignada__id__in=eMateriaAsignadaPlanificacionSedeVirtualExamenes,
                                    status=True),
                                                         nivel__periodo_id=periodo_id, status=True),
                                total_general=Count('materiaasignada__id',
                                                    filter=Q(nivel__periodo_id=periodo_id, status=True),
                                                    exclude=Q(materiaasignada__materia__asignatura__id=4837) | Q(
                                                        materiaasignada__materia__asignaturamalla__malla_id__in=eMallasIngles.values_list(
                                                            "id", flat=True)))).filter(
                                Q(total_general=F('total_planificadas')))
                            eMatriculas_exclude_planificadas_x_dia = eMatriculas.annotate(
                                total_planificadas=Count('materiaasignada__id', filter=Q(
                                    materiaasignada__id__in=eMateriaAsignadaPlanificacionSedeVirtualExamenes.filter(
                                        aulaplanificacion__turnoplanificacion__fechaplanificacion__fecha=fecha),
                                    status=True), nivel__periodo_id=periodo_id, status=True)).filter(
                                Q(total_planificadas=limite_x_día))
                            eMatriculas_exclude_ingles = eMatriculas.annotate(
                                total_ingles=Count('materiaasignada__materia__asignaturamalla__id', filter=Q(
                                    materiaasignada__materia__asignaturamalla__malla_id__in=eMallasIngles.values_list(
                                        "id",
                                        flat=True),
                                    nivel__periodo_id=periodo_id, status=True)),
                                total_general=Count('materiaasignada__materia__asignaturamalla__id',
                                                    filter=Q(nivel__periodo_id=periodo_id, status=True))).filter(
                                total_general=F('total_ingles'))
                            ids_exclude = list(
                                eMateriaAsignadaPlanificacionSedeVirtualExamenesmatriculas.filter(filter_conflicto))
                            ids_exclude.extend(list(eMatriculas_exclude_planificadas.values_list('id', flat=True)))
                            ids_exclude.extend(
                                list(eMatriculas_exclude_planificadas_x_dia.values_list('id', flat=True)))
                            ids_exclude.extend(list(eMatriculas_exclude_ingles.values_list('id', flat=True)))
                            sql = f"""SELECT
                                        "sga_matricula"."id",
                                        COUNT("sga_materia"."asignaturamalla_id")
                                                FILTER (WHERE ("sga_asignaturamalla"."malla_id" NOT IN (	SELECT U0."id"
                                                                                                                        FROM "sga_malla" U0
                                                                                                                        WHERE U0."id" IN (353, 22)
                                                                                                                    ) AND
                                                                    "sga_nivel"."periodo_id" = {periodo_id} AND
                                                                    "sga_matricula"."status"
                                                                    )
                                                        ) AS "total_general",
                                        COUNT("sga_materia"."asignaturamalla_id")
                                                FILTER (WHERE ("sga_materiaasignada"."id" IN (	SELECT U0."materiaasignada_id"
                                                                                                                        FROM "inno_materiaasignadaplanificacionsedevirtualexamen" U0
                                                                                                                        INNER JOIN "sga_materiaasignada" U1 ON U1."id" = U0.materiaasignada_id
                                                                                                                        WHERE U1."matricula_id" = "sga_matricula"."id"
                                                                                                                    ) AND
                                                                    "sga_nivel"."periodo_id" = {periodo_id} AND
                                                                    "sga_matricula"."status")
                                                        ) AS "total_planificadas"
                                    FROM "sga_matricula"
                                    INNER JOIN "sga_inscripcion" ON "sga_matricula"."inscripcion_id" = "sga_inscripcion"."id"
                                    INNER JOIN "sga_inscripcionmalla" ON "sga_inscripcion"."id" = "sga_inscripcionmalla"."inscripcion_id"
                                    INNER JOIN "sga_nivel" ON "sga_matricula"."nivel_id" = "sga_nivel"."id"
                                    INNER JOIN "sga_periodo" ON "sga_nivel"."periodo_id" = "sga_periodo"."id"
                                    LEFT OUTER JOIN "sga_materiaasignada" ON "sga_matricula"."id" = "sga_materiaasignada"."matricula_id"
                                    LEFT OUTER JOIN "sga_materia" ON "sga_materiaasignada"."materia_id" = "sga_materia"."id"
                                    LEFT OUTER JOIN "sga_asignaturamalla" ON "sga_materia"."asignaturamalla_id" = "sga_asignaturamalla"."id"
                                    WHERE (
                                        NOT "sga_matricula"."bloqueomatricula" AND
                                        "sga_inscripcionmalla"."malla_id" = {eMalla.pk} AND
                                        "sga_nivel"."periodo_id" = {periodo_id} AND
                                        "sga_matricula"."id" IN (
                                                                            SELECT DISTINCT
                                                                                U0."matricula_id"
                                                                            FROM "inno_matriculasedeexamen" U0
                                                                                INNER JOIN "sga_matricula" U2 ON U0."matricula_id" = U2."id"
                                                                                INNER JOIN "sga_nivel" U3 ON U2."nivel_id" = U3."id"
                                                                            WHERE (
                                                                                        U0."detallemodeloevaluativo_id" = {detallemodeloevaluativo_id} AND
                                                                                        U3."periodo_id" = {periodo_id} AND
                                                                                        NOT U2."retiradomatricula" AND
                                                                                        U2."status" AND
                                                                                        U0."sede_id" = {sede_id} AND
                                                                                        U0."status"
                                                                                    )
                                                                        ) AND
                                        NOT "sga_matricula"."retiradomatricula" AND
                                        "sga_matricula"."status"
                                        )
                                    GROUP BY "sga_matricula"."id"
                                    HAVING 
                                            COUNT("sga_materia"."asignaturamalla_id") 
                                                FILTER (WHERE ("sga_asignaturamalla"."malla_id" NOT IN (	SELECT U0."id"
                                                                                                                        FROM "sga_malla" U0
                                                                                                                        WHERE U0."id" IN (353, 22)
                                                                                                                    ) AND 
                                                                    "sga_nivel"."periodo_id" = {periodo_id} AND 
                                                                    "sga_matricula"."status"
                                                                    )
                                                        ) 
                                            <> 
                                            COUNT("sga_materia"."asignaturamalla_id") 
                                                FILTER (WHERE ("sga_materiaasignada"."id" IN (	SELECT U0."materiaasignada_id"
                                                                                                                        FROM "inno_materiaasignadaplanificacionsedevirtualexamen" U0
                                                                                                                        INNER JOIN "sga_materiaasignada" U1 ON U1."id" = U0.materiaasignada_id
                                                                                                                        WHERE U1."matricula_id" = "sga_matricula"."id"
                                                                                                                    ) AND
                                                                    "sga_nivel"."periodo_id" = {periodo_id} AND 
                                                                    "sga_matricula"."status")
                                    )"""
                            eMatriculas = eMatriculas.exclude(pk__in=ids_exclude)
                            cursor.execute(sql)
                            results = cursor.fetchall()
                            ids_matricula = [r[0] for r in results]
                            eMatriculas = eMatriculas.filter(pk__in=ids_matricula)
                            eMatriculas = eMatriculas.order_by('inscripcion__inscripcionnivel__nivel__orden',
                                                               'inscripcion__persona__apellido1',
                                                               'inscripcion__persona__apellido2',
                                                               'inscripcion__persona__nombres').distinct()
                            contador = cantidadad_planificadas
                            if not eMatriculas.values("id").exists():
                                contadorAulaSinLlenar += 1

                            if contadorAulaSinLlenar > 0:
                                # banderaBreakAula = True
                                break
                            for eMatricula in eMatriculas:
                                matricula_id = eMatricula.id
                                delet = MateriaAsignadaPlanificacionSedeVirtualExamen.objects.filter(status=True,
                                                                                                     materiaasignada__matricula_id=matricula_id).exclude(
                                    Q(materiaasignada__materia__asignatura__id=4837) |
                                    Q(materiaasignada__materia__asignaturamalla__malla_id__in=eMallasIngles.values_list(
                                        "id", flat=True))).values('id')
                                tmaterias = MateriaAsignada.objects.filter(status=True,
                                                                           matricula_id=matricula_id).exclude(
                                    Q(materia__asignatura__id=4837) |
                                    Q(materia__asignaturamalla__malla_id__in=eMallasIngles.values_list("id",
                                                                                                       flat=True))).values(
                                    'id')
                                if len(delet) == len(tmaterias) and not matricula_id in materias_completas:
                                    materias_completas.append(matricula_id)
                                eMateriaAsignadas = MateriaAsignada.objects.filter(status=True, matricula=eMatricula)
                                eMateriaAsignadas = eMateriaAsignadas.exclude(
                                    Q(pk__in=eMateriaAsignadaPlanificacionSedeVirtualExamenes) |
                                    Q(materia__asignatura__id=4837) |
                                    Q(materia__asignaturamalla__malla_id__in=eMallasIngles.values_list("id",
                                                                                                       flat=True)))
                                eMateriaAsignadas = eMateriaAsignadas.order_by(
                                    'materia__asignaturamalla__nivelmalla__orden')
                                if eMateriaAsignadas.values("id").exists():
                                    eMateriaAsignada = eMateriaAsignadas.first()
                                    contador += 1
                                    print(
                                        f"------- ({contador}) Se asignada el estudiante {eMateriaAsignada.matricula.inscripcion.persona} en la asignatura {eMateriaAsignada.materia.asignatura.nombre}")
                                    eMateriaAsignadaPlanificacionSedeVirtualExamen = MateriaAsignadaPlanificacionSedeVirtualExamen(
                                        aulaplanificacion=eAulaPlanificacionSedeVirtualExamen,
                                        materiaasignada=eMateriaAsignada,
                                        detallemodeloevaluativo_id=detallemodeloevaluativo_id)
                                    eMateriaAsignadaPlanificacionSedeVirtualExamen.save()
                                    eMateriaAsignada.visiblehorarioexamen = False
                                    eMateriaAsignada.save()
                                    if contador >= capacidad:
                                        break
    except Exception as ex:
        print(ex)


materias_no_planifica = [66002, 62844, 62823, 62845, 62946, 62944, 62945, 62947, 62943, 63022, 63020, 63021, 63023,
                         63019, 6363039,
                         63040, 63082, 63038, 63107, 63084, 63085, 63133, 63083, 63204, 63171, 63203, 63202, 63205,
                         63206, 65278, 65338, 65381, 63331, 63332, 63328,
                         63330, 63329, 63333, 63721, 64308, 64309, 64311, 64310, 64365, 64366, 64368, 64367, 64601,
                         64756, 64770, 64760, 64762, 64770, 64760, 64758, 63354]


def planificar_pregrado_unemi_v2_regulares(limite_x_día, sede, periodo_id, detallemodeloevaluativo_id, fechas):
    try:
        sede_id = sede
        print('PLANIFICANDO SEDE ' + str(sede_id))
        # periodo_id = 153
        # detallemodeloevaluativo_id = 37
        materias_completas = []
        cursor = connections['sga_select'].cursor()
        eMallasIngles = Malla.objects.filter(pk__in=[353, 22]).values_list('id', flat=True)
        eMallasMedicina = Malla.objects.filter(carrera__id__in=[223]).values_list('id', flat=True)
        eMatriculaSedeExamenes = MatriculaSedeExamen.objects.filter(sede_id=sede_id,
                                                                    detallemodeloevaluativo_id=detallemodeloevaluativo_id,
                                                                    status=True, matricula__status=True,
                                                                    matricula__retiradomatricula=False,
                                                                    matricula__nivel__periodo_id=periodo_id,
                                                                    matricula__inscripcion__carrera__modalidad=3).distinct()
        # if DEBUG:
        # MateriaAsignadaPlanificacionSedeVirtualExamen.objects.filter(status=True,
        #                                                              aulaplanificacion__turnoplanificacion__fechaplanificacion__periodo_id=periodo_id,
        #                                                              aulaplanificacion__turnoplanificacion__fechaplanificacion__sede_id=sede_id,
        #                                                              detallemodeloevaluativo_id=detallemodeloevaluativo_id).delete()

        eMatriculas = Matricula.objects.filter(pk__in=eMatriculaSedeExamenes.values_list("matricula__id", flat=True),
                                               status=True, retiradomatricula=False, bloqueomatricula=False,
                                               nivel__periodo_id=periodo_id)
        eMatriculas_exclude_ingles = eMatriculas.annotate(
            total_ingles=Count('materiaasignada__materia__asignaturamalla__id', filter=Q(
                materiaasignada__materia__asignaturamalla__malla_id__in=eMallasIngles.values_list('id', flat=True),
                nivel__periodo_id=periodo_id, status=True)),
            total_general=Count('materiaasignada__materia__asignaturamalla__id',
                                filter=Q(nivel__periodo_id=periodo_id, status=True))).filter(
            total_general=F('total_ingles'))
        ids_exclude = []
        ids_exclude.extend(list(eMatriculas_exclude_ingles.values_list('id', flat=True)))
        eMatriculas = eMatriculas.exclude(pk__in=ids_exclude)
        eMallas = Malla.objects.filter(carrera__modalidad=3,
                                       pk__in=eMatriculas.values_list('inscripcion__inscripcionmalla__malla_id',
                                                                      flat=True).distinct()).exclude(
            id__in=eMallasIngles)
        # eCarreras = Carrera.objects.filter(pk__in=eMallas.values_list('carrera__id', flat=True))
        # eCoordinaciones = Coordinacion.objects.filter(pk__in=eCarreras.values_list('coordinacion__id', flat=True)).distinct()
        # for eCoordinacion in eCoordinaciones:
        #     eMallas = eMallas.filter(carrera_id__in=eCoordinacion.carreras().values_list('id', flat=True)).order_by('carrera__nombre', 'inicio')
        eMallas = eMallas.order_by('carrera__nombre', 'inicio')
        for eMalla in eMallas:
            eFechaPlanificacionSedeVirtualExamenes = FechaPlanificacionSedeVirtualExamen.objects.filter(sede_id=sede_id,
                                                                                                        periodo_id=periodo_id,
                                                                                                        fecha__in=fechas).order_by(
                'fecha')
            for eFechaPlanificacionSedeVirtualExamen in eFechaPlanificacionSedeVirtualExamenes:
                fecha = eFechaPlanificacionSedeVirtualExamen.fecha
                eTurnoPlanificacionSedeVirtualExamenes = TurnoPlanificacionSedeVirtualExamen.objects.filter(
                    fechaplanificacion=eFechaPlanificacionSedeVirtualExamen).order_by('horainicio')
                for eTurnoPlanificacionSedeVirtualExamen in eTurnoPlanificacionSedeVirtualExamenes:
                    horainicio = eTurnoPlanificacionSedeVirtualExamen.horainicio
                    horafin = eTurnoPlanificacionSedeVirtualExamen.horafin

                    # eAulaPlanificacionSedeVirtualExamenes = AulaPlanificacionSedeVirtualExamen.objects.filter(turnoplanificacion__fechaplanificacion__sede_id=sede_id,
                    #                                                                                           turnoplanificacion__fechaplanificacion__periodo_id=periodo_id)
                    eAulaPlanificacionSedeVirtualExamenes = AulaPlanificacionSedeVirtualExamen.objects.filter(
                        turnoplanificacion=eTurnoPlanificacionSedeVirtualExamen)
                    eAulaPlanificacionSedeVirtualExamenes_exclude_llenos = eAulaPlanificacionSedeVirtualExamenes.annotate(
                        total_general=Count('materiaasignadaplanificacionsedevirtualexamen__id', filter=Q(
                            materiaasignadaplanificacionsedevirtualexamen__materiaasignada__matricula__nivel__periodo_id=periodo_id,
                            status=True))).filter(total_general=F('aula__capacidad'))
                    eAulaPlanificacionSedeVirtualExamenes.exclude(
                        pk__in=eAulaPlanificacionSedeVirtualExamenes_exclude_llenos.values_list("id", flat=True))
                    eAulaPlanificacionSedeVirtualExamenes = eAulaPlanificacionSedeVirtualExamenes.order_by(
                        'turnoplanificacion__fechaplanificacion__fecha',
                        'turnoplanificacion__horainicio').distinct()
                    # totalAulaSinLlenar = len(eAulaPlanificacionSedeVirtualExamenes.values("id"))
                    contadorAulaSinLlenar = 0
                    # banderaBreakAula = False
                    for eAulaPlanificacionSedeVirtualExamen in eAulaPlanificacionSedeVirtualExamenes:
                        eAula = eAulaPlanificacionSedeVirtualExamen.aula
                        capacidad = eAula.capacidad
                        cantidadad_planificadas = eAulaPlanificacionSedeVirtualExamen.cantidadad_planificadas()
                        eTurnoPlanificacionSedeVirtualExamen = eAulaPlanificacionSedeVirtualExamen.turnoplanificacion
                        # horainicio = eTurnoPlanificacionSedeVirtualExamen.horainicio
                        # horafin = eTurnoPlanificacionSedeVirtualExamen.horafin
                        eFechaPlanificacionSedeVirtualExamen = eTurnoPlanificacionSedeVirtualExamen.fechaplanificacion
                        # fecha = eFechaPlanificacionSedeVirtualExamen.fecha
                        if cantidadad_planificadas < capacidad:
                            print(
                                f"Se procede a planificar en la fecha {fecha} en el horario {horainicio} a {horafin} en el aula {eAula.nombre}")
                            eMateriaAsignadaPlanificacionSedeVirtualExamenes = MateriaAsignadaPlanificacionSedeVirtualExamen.objects.filter(
                                status=True,
                                aulaplanificacion__turnoplanificacion__fechaplanificacion__periodo_id=periodo_id,
                                detallemodeloevaluativo_id=detallemodeloevaluativo_id).exclude(
                                materiaasignada__materia_id__in=materias_no_planifica).values_list(
                                "materiaasignada__id",
                                flat=True)
                            eMateriaAsignadaPlanificacionSedeVirtualExamenesmatriculas = MateriaAsignadaPlanificacionSedeVirtualExamen.objects.filter(
                                status=True,
                                aulaplanificacion__turnoplanificacion__fechaplanificacion__periodo_id=periodo_id,
                                detallemodeloevaluativo_id=detallemodeloevaluativo_id).exclude(
                                materiaasignada__materia_id__in=materias_no_planifica).values_list(
                                "materiaasignada__matricula__id", flat=True)

                            filter_conflicto = (Q(aulaplanificacion__turnoplanificacion__horainicio__lte=horafin,
                                                  aulaplanificacion__turnoplanificacion__horafin__gte=horafin,
                                                  aulaplanificacion__turnoplanificacion__fechaplanificacion__fecha=fecha) |
                                                Q(aulaplanificacion__turnoplanificacion__horainicio__lte=horainicio,
                                                  aulaplanificacion__turnoplanificacion__horafin__gte=horainicio,
                                                  aulaplanificacion__turnoplanificacion__fechaplanificacion__fecha=fecha))
                            eMatriculas = Matricula.objects.filter(
                                pk__in=eMatriculaSedeExamenes.values_list("matricula__id", flat=True),
                                status=True,
                                retiradomatricula=False,
                                bloqueomatricula=False,
                                nivel__periodo_id=periodo_id,
                                inscripcion__inscripcionmalla__malla=eMalla).exclude(id__in=materias_completas)
                            # eMatriculas = eMatriculas.filter(pk__in=ids_matricula)
                            eMatriculas_exclude_planificadas = eMatriculas.annotate(
                                total_planificadas=Count('materiaasignada__id', filter=Q(
                                    materiaasignada__id__in=eMateriaAsignadaPlanificacionSedeVirtualExamenes,
                                    status=True),
                                                         nivel__periodo_id=periodo_id, status=True),
                                total_general=Count('materiaasignada__id',
                                                    filter=Q(nivel__periodo_id=periodo_id, status=True),
                                                    exclude=Q(materiaasignada__materia__asignatura__id=4837) | Q(
                                                        materiaasignada__materia__asignaturamalla__malla_id__in=eMallasIngles.values_list(
                                                            "id", flat=True)))).filter(
                                Q(total_general=F('total_planificadas')))
                            eMatriculas_exclude_planificadas_x_dia = eMatriculas.annotate(
                                total_planificadas=Count('materiaasignada__id', filter=Q(
                                    materiaasignada__id__in=eMateriaAsignadaPlanificacionSedeVirtualExamenes.filter(
                                        aulaplanificacion__turnoplanificacion__fechaplanificacion__fecha=fecha),
                                    status=True), nivel__periodo_id=periodo_id, status=True)).filter(
                                Q(total_planificadas=limite_x_día))
                            eMatriculas_exclude_ingles = eMatriculas.annotate(
                                total_ingles=Count('materiaasignada__materia__asignaturamalla__id', filter=Q(
                                    materiaasignada__materia__asignaturamalla__malla_id__in=eMallasIngles.values_list(
                                        "id",
                                        flat=True),
                                    nivel__periodo_id=periodo_id, status=True)),
                                total_general=Count('materiaasignada__materia__asignaturamalla__id',
                                                    filter=Q(nivel__periodo_id=periodo_id, status=True))).filter(
                                total_general=F('total_ingles'))
                            ids_exclude = list(
                                eMateriaAsignadaPlanificacionSedeVirtualExamenesmatriculas.filter(filter_conflicto))
                            ids_exclude.extend(list(eMatriculas_exclude_planificadas.values_list('id', flat=True)))
                            ids_exclude.extend(
                                list(eMatriculas_exclude_planificadas_x_dia.values_list('id', flat=True)))
                            ids_exclude.extend(list(eMatriculas_exclude_ingles.values_list('id', flat=True)))
                            sql = f"""SELECT 
                                        "sga_matricula"."id", 
                                        COUNT("sga_materia"."asignaturamalla_id") 
                                                FILTER (WHERE ("sga_asignaturamalla"."malla_id" NOT IN (	SELECT U0."id"
                                                                                                                        FROM "sga_malla" U0
                                                                                                                        WHERE U0."id" IN (353, 22)
                                                                                                                    ) AND 
                                                                    "sga_nivel"."periodo_id" = {periodo_id} AND 
                                                                    "sga_matricula"."status"
                                                                    )
                                                        ) AS "total_general", 
                                        COUNT("sga_materia"."asignaturamalla_id") 
                                                FILTER (WHERE ("sga_materiaasignada"."id" IN (	SELECT U0."materiaasignada_id"
                                                                                                                        FROM "inno_materiaasignadaplanificacionsedevirtualexamen" U0
                                                                                                                        INNER JOIN "sga_materiaasignada" U1 ON U1."id" = U0.materiaasignada_id
                                                                                                                        WHERE U1."matricula_id" = "sga_matricula"."id"
                                                                                                                    ) AND
                                                                    "sga_nivel"."periodo_id" = {periodo_id} AND 
                                                                    "sga_matricula"."status")
                                                        ) AS "total_planificadas"
                                    FROM "sga_matricula"
                                    INNER JOIN "sga_inscripcion" ON "sga_matricula"."inscripcion_id" = "sga_inscripcion"."id"
                                    INNER JOIN "sga_inscripcionmalla" ON "sga_inscripcion"."id" = "sga_inscripcionmalla"."inscripcion_id"
                                    INNER JOIN "sga_nivel" ON "sga_matricula"."nivel_id" = "sga_nivel"."id"
                                    INNER JOIN "sga_periodo" ON "sga_nivel"."periodo_id" = "sga_periodo"."id"
                                    LEFT OUTER JOIN "sga_materiaasignada" ON "sga_matricula"."id" = "sga_materiaasignada"."matricula_id"
                                    LEFT OUTER JOIN "sga_materia" ON "sga_materiaasignada"."materia_id" = "sga_materia"."id"
                                    LEFT OUTER JOIN "sga_asignaturamalla" ON "sga_materia"."asignaturamalla_id" = "sga_asignaturamalla"."id"
                                    WHERE (
                                        NOT "sga_matricula"."bloqueomatricula" AND 
                                        "sga_inscripcionmalla"."malla_id" = {eMalla.pk} AND 
                                        "sga_nivel"."periodo_id" = {periodo_id} AND 
                                        "sga_matricula"."id" IN (
                                                                            SELECT DISTINCT 
                                                                                U0."matricula_id"
                                                                            FROM "inno_matriculasedeexamen" U0
                                                                                INNER JOIN "sga_matricula" U2 ON U0."matricula_id" = U2."id"
                                                                                INNER JOIN "sga_nivel" U3 ON U2."nivel_id" = U3."id"
                                                                            WHERE (
                                                                                        U0."detallemodeloevaluativo_id" = {detallemodeloevaluativo_id} AND 
                                                                                        U3."periodo_id" = {periodo_id} AND 
                                                                                        NOT U2."retiradomatricula" AND 
                                                                                        U2."status" AND 
                                                                                        U0."sede_id" = {sede_id} AND 
                                                                                        U0."status"
                                                                                    )
                                                                        ) AND 
                                        NOT "sga_matricula"."retiradomatricula" AND 
                                        "sga_matricula"."status"
                                        )
                                    GROUP BY "sga_matricula"."id"
                                    HAVING 
                                            COUNT("sga_materia"."asignaturamalla_id") 
                                                FILTER (WHERE ("sga_asignaturamalla"."malla_id" NOT IN (	SELECT U0."id"
                                                                                                                        FROM "sga_malla" U0
                                                                                                                        WHERE U0."id" IN (353, 22)
                                                                                                                    ) AND 
                                                                    "sga_nivel"."periodo_id" = {periodo_id} AND 
                                                                    "sga_matricula"."status"
                                                                    )
                                                        ) 
                                            <> 
                                            COUNT("sga_materia"."asignaturamalla_id") 
                                                FILTER (WHERE ("sga_materiaasignada"."id" IN (	SELECT U0."materiaasignada_id"
                                                                                                                        FROM "inno_materiaasignadaplanificacionsedevirtualexamen" U0
                                                                                                                        INNER JOIN "sga_materiaasignada" U1 ON U1."id" = U0.materiaasignada_id
                                                                                                                        WHERE U1."matricula_id" = "sga_matricula"."id"
                                                                                                                    ) AND
                                                                    "sga_nivel"."periodo_id" = {periodo_id} AND 
                                                                    "sga_matricula"."status")
                                    )"""
                            eMatriculas = eMatriculas.exclude(pk__in=ids_exclude)
                            cursor.execute(sql)
                            results = cursor.fetchall()
                            ids_matricula = [r[0] for r in results]
                            eMatriculas = eMatriculas.filter(pk__in=ids_matricula)
                            eMatriculas = eMatriculas.order_by('inscripcion__inscripcionnivel__nivel__orden',
                                                               'inscripcion__persona__apellido1',
                                                               'inscripcion__persona__apellido2',
                                                               'inscripcion__persona__nombres').distinct()
                            contador = cantidadad_planificadas
                            if not eMatriculas.values("id").exists():
                                contadorAulaSinLlenar += 1

                            if contadorAulaSinLlenar > 0:
                                # banderaBreakAula = True
                                break
                            for eMatricula in eMatriculas:
                                matricula_id = eMatricula.id
                                mat = MatriculaSedeExamen.objects.filter(status=True, matricula=eMatricula,
                                                                         sede_id=sede_id,
                                                                         detallemodeloevaluativo_id=detallemodeloevaluativo_id).first()
                                delet = MateriaAsignadaPlanificacionSedeVirtualExamen.objects.filter(status=True,
                                                                                                     materiaasignada__matricula_id=matricula_id).exclude(
                                    Q(materiaasignada__materia__asignatura__id=4837) |
                                    Q(materiaasignada__materia__asignaturamalla__malla_id__in=eMallasIngles.values_list(
                                        "id", flat=True))).exclude(
                                    materiaasignada__materia_id__in=materias_no_planifica).values('id')
                                tmaterias = MateriaAsignada.objects.filter(status=True,
                                                                           matricula_id=matricula_id).exclude(
                                    Q(materia__asignatura__id=4837) |
                                    Q(materia__asignaturamalla__malla_id__in=eMallasIngles.values_list("id",
                                                                                                       flat=True))).exclude(
                                    materia_id__in=materias_no_planifica).values(
                                    'id')
                                # if len(tmaterias) < 5 or len(tmaterias) > 6:
                                #     continue
                                if len(delet) == len(tmaterias) and not matricula_id in materias_completas:
                                    materias_completas.append(matricula_id)
                                eMateriaAsignadas = MateriaAsignada.objects.filter(status=True,
                                                                                   matricula=eMatricula).exclude(
                                    materia_id__in=materias_no_planifica)
                                eMateriaAsignadas = eMateriaAsignadas.exclude(
                                    Q(pk__in=eMateriaAsignadaPlanificacionSedeVirtualExamenes) |
                                    Q(materia__asignatura__id=4837) |
                                    Q(materia__asignaturamalla__malla_id__in=eMallasIngles.values_list("id",
                                                                                                       flat=True)))
                                eMateriaAsignadas = eMateriaAsignadas.order_by(
                                    'materia__asignaturamalla__nivelmalla__orden')
                                if eMateriaAsignadas.values("id").exists():
                                    eMateriaAsignada = eMateriaAsignadas.first()
                                    contador += 1
                                    print(
                                        f"------- ({contador}) Se asignada el estudiante {eMateriaAsignada.matricula.inscripcion.persona} en la asignatura {eMateriaAsignada.materia.asignatura.nombre}")
                                    eMateriaAsignadaPlanificacionSedeVirtualExamen = MateriaAsignadaPlanificacionSedeVirtualExamen(
                                        aulaplanificacion=eAulaPlanificacionSedeVirtualExamen,
                                        materiaasignada=eMateriaAsignada,
                                        detallemodeloevaluativo_id=detallemodeloevaluativo_id,
                                        utilizar_qr=True
                                    )
                                    eMateriaAsignadaPlanificacionSedeVirtualExamen.save()
                                    if mat:
                                        if not mat.aceptotermino:
                                            result = generate_qr_examen_final(
                                                eMateriaAsignadaPlanificacionSedeVirtualExamen,
                                                materiaasignada_id=eMateriaAsignadaPlanificacionSedeVirtualExamen.materiaasignada_id)

                                    eMateriaAsignada.save()
                                    for mate in MateriaAsignada.objects.filter(status=True, matricula=eMatricula):
                                        mate.visiblehorarioexamen = False
                                        mate.save()
                                    if contador >= capacidad:
                                        break
    except Exception as ex:
        print(ex)


def planificar_pregrado_unemi_virtual_v2(limite_x_día, sede, periodo_id, detallemodeloevaluativo_id, fechas):
    try:
        sede_id = sede
        # periodo_id = 153
        # detallemodeloevaluativo_id = 37
        materias_completas = []
        cursor = connections['sga_select'].cursor()
        eMallasIngles = Malla.objects.filter(pk__in=[353, 22]).values_list('id', flat=True)
        eMallasMedicina = Malla.objects.filter(carrera__id__in=[223]).values_list('id', flat=True)
        eMatriculaSedeExamenes = MatriculaSedeExamen.objects.filter(sede_id=sede_id,
                                                                    detallemodeloevaluativo_id=detallemodeloevaluativo_id,
                                                                    status=True, matricula__status=True,
                                                                    matricula__retiradomatricula=False,
                                                                    matricula__nivel__periodo_id=periodo_id).distinct()
        # if DEBUG:
        # MateriaAsignadaPlanificacionSedeVirtualExamen.objects.filter(status=True,
        #                                                              aulaplanificacion__turnoplanificacion__fechaplanificacion__periodo_id=periodo_id,
        #                                                              aulaplanificacion__turnoplanificacion__fechaplanificacion__sede_id=sede_id,
        #                                                              detallemodeloevaluativo_id=detallemodeloevaluativo_id).delete()

        eMatriculas = Matricula.objects.filter(pk__in=eMatriculaSedeExamenes.values_list("matricula__id", flat=True),
                                               status=True, retiradomatricula=False, bloqueomatricula=False,
                                               nivel__periodo_id=periodo_id)
        eMatriculas_exclude_ingles = eMatriculas.annotate(
            total_ingles=Count('materiaasignada__materia__asignaturamalla__id', filter=Q(
                materiaasignada__materia__asignaturamalla__malla_id__in=eMallasIngles.values_list('id', flat=True),
                nivel__periodo_id=periodo_id, status=True)),
            total_general=Count('materiaasignada__materia__asignaturamalla__id',
                                filter=Q(nivel__periodo_id=periodo_id, status=True))).filter(
            total_general=F('total_ingles'))
        ids_exclude = []
        ids_exclude.extend(list(eMatriculas_exclude_ingles.values_list('id', flat=True)))
        eMatriculas = eMatriculas.exclude(pk__in=ids_exclude)
        eMallas = Malla.objects.filter(
            pk__in=eMatriculas.values_list('inscripcion__inscripcionmalla__malla_id', flat=True).distinct()).exclude(
            id__in=eMallasIngles)
        # eCarreras = Carrera.objects.filter(pk__in=eMallas.values_list('carrera__id', flat=True))
        # eCoordinaciones = Coordinacion.objects.filter(pk__in=eCarreras.values_list('coordinacion__id', flat=True)).distinct()
        # for eCoordinacion in eCoordinaciones:
        #     eMallas = eMallas.filter(carrera_id__in=eCoordinacion.carreras().values_list('id', flat=True)).order_by('carrera__nombre', 'inicio')
        eMallas = eMallas.order_by('carrera__nombre', 'inicio')
        for eMalla in eMallas:
            eFechaPlanificacionSedeVirtualExamenes = FechaPlanificacionSedeVirtualExamen.objects.filter(sede_id=sede_id,
                                                                                                        periodo_id=periodo_id,
                                                                                                        fecha__in=fechas).order_by(
                'fecha')
            for eFechaPlanificacionSedeVirtualExamen in eFechaPlanificacionSedeVirtualExamenes:
                fecha = eFechaPlanificacionSedeVirtualExamen.fecha
                eTurnoPlanificacionSedeVirtualExamenes = TurnoPlanificacionSedeVirtualExamen.objects.filter(
                    fechaplanificacion=eFechaPlanificacionSedeVirtualExamen).order_by('horainicio')
                for eTurnoPlanificacionSedeVirtualExamen in eTurnoPlanificacionSedeVirtualExamenes:
                    horainicio = eTurnoPlanificacionSedeVirtualExamen.horainicio
                    horafin = eTurnoPlanificacionSedeVirtualExamen.horafin

                    # eAulaPlanificacionSedeVirtualExamenes = AulaPlanificacionSedeVirtualExamen.objects.filter(turnoplanificacion__fechaplanificacion__sede_id=sede_id,
                    #                                                                                           turnoplanificacion__fechaplanificacion__periodo_id=periodo_id)
                    eAulaPlanificacionSedeVirtualExamenes = AulaPlanificacionSedeVirtualExamen.objects.filter(
                        turnoplanificacion=eTurnoPlanificacionSedeVirtualExamen)
                    eAulaPlanificacionSedeVirtualExamenes_exclude_llenos = eAulaPlanificacionSedeVirtualExamenes.annotate(
                        total_general=Count('materiaasignadaplanificacionsedevirtualexamen__id', filter=Q(
                            materiaasignadaplanificacionsedevirtualexamen__materiaasignada__matricula__nivel__periodo_id=periodo_id,
                            status=True))).filter(total_general=F('aula__capacidad'))
                    eAulaPlanificacionSedeVirtualExamenes.exclude(
                        pk__in=eAulaPlanificacionSedeVirtualExamenes_exclude_llenos.values_list("id", flat=True))
                    eAulaPlanificacionSedeVirtualExamenes = eAulaPlanificacionSedeVirtualExamenes.order_by(
                        'turnoplanificacion__fechaplanificacion__fecha',
                        'turnoplanificacion__horainicio').distinct()
                    # totalAulaSinLlenar = len(eAulaPlanificacionSedeVirtualExamenes.values("id"))
                    contadorAulaSinLlenar = 0
                    # banderaBreakAula = False
                    for eAulaPlanificacionSedeVirtualExamen in eAulaPlanificacionSedeVirtualExamenes:
                        eAula = eAulaPlanificacionSedeVirtualExamen.aula
                        capacidad = eAula.capacidad
                        cantidadad_planificadas = eAulaPlanificacionSedeVirtualExamen.cantidadad_planificadas()
                        eTurnoPlanificacionSedeVirtualExamen = eAulaPlanificacionSedeVirtualExamen.turnoplanificacion
                        # horainicio = eTurnoPlanificacionSedeVirtualExamen.horainicio
                        # horafin = eTurnoPlanificacionSedeVirtualExamen.horafin
                        eFechaPlanificacionSedeVirtualExamen = eTurnoPlanificacionSedeVirtualExamen.fechaplanificacion
                        # fecha = eFechaPlanificacionSedeVirtualExamen.fecha
                        if cantidadad_planificadas < capacidad:
                            print(
                                f"Se procede a planificar en la fecha {fecha} en el horario {horainicio} a {horafin} en el aula {eAula.nombre}")
                            eMateriaAsignadaPlanificacionSedeVirtualExamenes = MateriaAsignadaPlanificacionSedeVirtualExamen.objects.filter(
                                status=True,
                                aulaplanificacion__turnoplanificacion__fechaplanificacion__periodo_id=periodo_id,
                                detallemodeloevaluativo_id=detallemodeloevaluativo_id).exclude(
                                materiaasignada__materia_id__in=materias_no_planifica).values_list(
                                "materiaasignada__id",
                                flat=True)
                            eMateriaAsignadaPlanificacionSedeVirtualExamenesmatriculas = MateriaAsignadaPlanificacionSedeVirtualExamen.objects.filter(
                                status=True,
                                aulaplanificacion__turnoplanificacion__fechaplanificacion__periodo_id=periodo_id,
                                detallemodeloevaluativo_id=detallemodeloevaluativo_id).values_list(
                                "materiaasignada__matricula__id", flat=True)

                            filter_conflicto = []
                            eMatriculas = Matricula.objects.filter(
                                pk__in=eMatriculaSedeExamenes.values_list("matricula__id", flat=True),
                                status=True,
                                retiradomatricula=False,
                                bloqueomatricula=False,
                                nivel__periodo_id=periodo_id,
                                inscripcion__inscripcionmalla__malla=eMalla).exclude(id__in=materias_completas)
                            # eMatriculas = eMatriculas.filter(pk__in=ids_matricula)
                            eMatriculas_exclude_planificadas = eMatriculas.annotate(
                                total_planificadas=Count('materiaasignada__id', filter=Q(
                                    materiaasignada__id__in=eMateriaAsignadaPlanificacionSedeVirtualExamenes,
                                    status=True),
                                                         nivel__periodo_id=periodo_id, status=True),
                                total_general=Count('materiaasignada__id',
                                                    filter=Q(nivel__periodo_id=periodo_id, status=True),
                                                    exclude=Q(materiaasignada__materia__asignatura__id=4837) | Q(
                                                        materiaasignada__materia__asignaturamalla__malla_id__in=eMallasIngles.values_list(
                                                            "id", flat=True)))).filter(
                                Q(total_general=F('total_planificadas')))
                            eMatriculas_exclude_planificadas_x_dia = eMatriculas.annotate(
                                total_planificadas=Count('materiaasignada__id', filter=Q(
                                    materiaasignada__id__in=eMateriaAsignadaPlanificacionSedeVirtualExamenes.filter(
                                        aulaplanificacion__turnoplanificacion__fechaplanificacion__fecha=fecha),
                                    status=True), nivel__periodo_id=periodo_id, status=True)).filter(
                                Q(total_planificadas=limite_x_día))
                            eMatriculas_exclude_ingles = eMatriculas.annotate(
                                total_ingles=Count('materiaasignada__materia__asignaturamalla__id', filter=Q(
                                    materiaasignada__materia__asignaturamalla__malla_id__in=eMallasIngles.values_list(
                                        "id",
                                        flat=True),
                                    nivel__periodo_id=periodo_id, status=True)),
                                total_general=Count('materiaasignada__materia__asignaturamalla__id',
                                                    filter=Q(nivel__periodo_id=periodo_id, status=True))).filter(
                                total_general=F('total_ingles'))
                            ids_exclude.extend(list(eMatriculas_exclude_planificadas.values_list('id', flat=True)))
                            ids_exclude.extend(
                                list(eMatriculas_exclude_planificadas_x_dia.values_list('id', flat=True)))
                            ids_exclude.extend(list(eMatriculas_exclude_ingles.values_list('id', flat=True)))
                            sql = f"""SELECT 
                                        "sga_matricula"."id", 
                                        COUNT("sga_materia"."asignaturamalla_id") 
                                                FILTER (WHERE ("sga_asignaturamalla"."malla_id" NOT IN (	SELECT U0."id"
                                                                                                                        FROM "sga_malla" U0
                                                                                                                        WHERE U0."id" IN (353, 22)
                                                                                                                    ) AND 
                                                                    "sga_nivel"."periodo_id" = {periodo_id} AND 
                                                                    "sga_matricula"."status"
                                                                    )
                                                        ) AS "total_general", 
                                        COUNT("sga_materia"."asignaturamalla_id") 
                                                FILTER (WHERE ("sga_materiaasignada"."id" IN (	SELECT U0."materiaasignada_id"
                                                                                                                        FROM "inno_materiaasignadaplanificacionsedevirtualexamen" U0
                                                                                                                        INNER JOIN "sga_materiaasignada" U1 ON U1."id" = U0.materiaasignada_id
                                                                                                                        WHERE U1."matricula_id" = "sga_matricula"."id"
                                                                                                                    ) AND
                                                                    "sga_nivel"."periodo_id" = {periodo_id} AND 
                                                                    "sga_matricula"."status")
                                                        ) AS "total_planificadas"
                                    FROM "sga_matricula"
                                    INNER JOIN "sga_inscripcion" ON "sga_matricula"."inscripcion_id" = "sga_inscripcion"."id"
                                    INNER JOIN "sga_inscripcionmalla" ON "sga_inscripcion"."id" = "sga_inscripcionmalla"."inscripcion_id"
                                    INNER JOIN "sga_nivel" ON "sga_matricula"."nivel_id" = "sga_nivel"."id"
                                    INNER JOIN "sga_periodo" ON "sga_nivel"."periodo_id" = "sga_periodo"."id"
                                    LEFT OUTER JOIN "sga_materiaasignada" ON "sga_matricula"."id" = "sga_materiaasignada"."matricula_id"
                                    LEFT OUTER JOIN "sga_materia" ON "sga_materiaasignada"."materia_id" = "sga_materia"."id"
                                    LEFT OUTER JOIN "sga_asignaturamalla" ON "sga_materia"."asignaturamalla_id" = "sga_asignaturamalla"."id"
                                    WHERE (
                                        NOT "sga_matricula"."bloqueomatricula" AND 
                                        "sga_inscripcionmalla"."malla_id" = {eMalla.pk} AND 
                                        "sga_nivel"."periodo_id" = {periodo_id} AND 
                                        "sga_matricula"."id" IN (
                                                                            SELECT DISTINCT 
                                                                                U0."matricula_id"
                                                                            FROM "inno_matriculasedeexamen" U0
                                                                                INNER JOIN "sga_matricula" U2 ON U0."matricula_id" = U2."id"
                                                                                INNER JOIN "sga_nivel" U3 ON U2."nivel_id" = U3."id"
                                                                            WHERE (
                                                                                        U0."detallemodeloevaluativo_id" = {detallemodeloevaluativo_id} AND 
                                                                                        U3."periodo_id" = {periodo_id} AND 
                                                                                        NOT U2."retiradomatricula" AND 
                                                                                        U2."status" AND 
                                                                                        U0."sede_id" = {sede_id} AND 
                                                                                        U0."status"
                                                                                    )
                                                                        ) AND 
                                        NOT "sga_matricula"."retiradomatricula" AND 
                                        "sga_matricula"."status"
                                        )
                                    GROUP BY "sga_matricula"."id"
                                    HAVING 
                                            COUNT("sga_materia"."asignaturamalla_id") 
                                                FILTER (WHERE ("sga_asignaturamalla"."malla_id" NOT IN (	SELECT U0."id"
                                                                                                                        FROM "sga_malla" U0
                                                                                                                        WHERE U0."id" IN (353, 22)
                                                                                                                    ) AND 
                                                                    "sga_nivel"."periodo_id" = {periodo_id} AND 
                                                                    "sga_matricula"."status"
                                                                    )
                                                        ) 
                                            <> 
                                            COUNT("sga_materia"."asignaturamalla_id") 
                                                FILTER (WHERE ("sga_materiaasignada"."id" IN (	SELECT U0."materiaasignada_id"
                                                                                                                        FROM "inno_materiaasignadaplanificacionsedevirtualexamen" U0
                                                                                                                        INNER JOIN "sga_materiaasignada" U1 ON U1."id" = U0.materiaasignada_id
                                                                                                                        WHERE U1."matricula_id" = "sga_matricula"."id"
                                                                                                                    ) AND
                                                                    "sga_nivel"."periodo_id" = {periodo_id} AND 
                                                                    "sga_matricula"."status")
                                    )"""
                            eMatriculas = eMatriculas.exclude(pk__in=ids_exclude)
                            cursor.execute(sql)
                            results = cursor.fetchall()
                            ids_matricula = [r[0] for r in results]
                            eMatriculas = eMatriculas.filter(pk__in=ids_matricula)
                            eMatriculas = eMatriculas.order_by('inscripcion__inscripcionnivel__nivel__orden',
                                                               'inscripcion__persona__apellido1',
                                                               'inscripcion__persona__apellido2',
                                                               'inscripcion__persona__nombres').distinct()
                            contador = cantidadad_planificadas
                            if not eMatriculas.values("id").exists():
                                contadorAulaSinLlenar += 1

                            if contadorAulaSinLlenar > 0:
                                # banderaBreakAula = True
                                break
                            for eMatricula in eMatriculas:
                                matricula_id = eMatricula.id
                                delet = MateriaAsignadaPlanificacionSedeVirtualExamen.objects.filter(status=True,
                                                                                                     materiaasignada__matricula_id=matricula_id).exclude(
                                    Q(materiaasignada__materia__asignatura__id=4837) |
                                    Q(materiaasignada__materia__asignaturamalla__malla_id__in=eMallasIngles.values_list(
                                        "id", flat=True))).exclude(
                                    materiaasignada__materia_id__in=materias_no_planifica).values('id')
                                tmaterias = MateriaAsignada.objects.filter(status=True,
                                                                           matricula_id=matricula_id).exclude(
                                    Q(materia__asignatura__id=4837) |
                                    Q(materia__asignaturamalla__malla_id__in=eMallasIngles.values_list("id",
                                                                                                       flat=True))).exclude(
                                    materia_id__in=materias_no_planifica).values(
                                    'id')
                                if len(delet) == len(tmaterias) and not matricula_id in materias_completas:
                                    materias_completas.append(matricula_id)
                                eMateriaAsignadas = MateriaAsignada.objects.filter(status=True, matricula=eMatricula)

                                eMateriaAsignadas = eMateriaAsignadas.exclude(
                                    materia_id__in=materias_no_planifica).exclude(
                                    Q(pk__in=eMateriaAsignadaPlanificacionSedeVirtualExamenes) |
                                    Q(materia__asignatura__id=4837) |
                                    Q(materia__asignaturamalla__malla_id__in=eMallasIngles.values_list("id",
                                                                                                       flat=True)))
                                eMateriaAsignadas = eMateriaAsignadas.order_by(
                                    'materia__asignaturamalla__nivelmalla__orden')
                                if eMateriaAsignadas.values("id").exists():
                                    eMateriaAsignada = eMateriaAsignadas.first()
                                    if MateriaAsignadaPlanificacionSedeVirtualExamen.objects.filter(
                                            aulaplanificacion=eAulaPlanificacionSedeVirtualExamen,
                                            materiaasignada=eMateriaAsignada,
                                            detallemodeloevaluativo_id=detallemodeloevaluativo_id,
                                            utilizar_qr=True).exists():
                                        continue
                                    contador += 1
                                    print(
                                        f"------- ({contador}) Se asignada el estudiante {eMateriaAsignada.matricula.inscripcion.persona} en la asignatura {eMateriaAsignada.materia.asignatura.nombre}")
                                    eMateriaAsignadaPlanificacionSedeVirtualExamen = MateriaAsignadaPlanificacionSedeVirtualExamen(
                                        aulaplanificacion=eAulaPlanificacionSedeVirtualExamen,
                                        materiaasignada=eMateriaAsignada,
                                        detallemodeloevaluativo_id=detallemodeloevaluativo_id,
                                        utilizar_qr=True)
                                    eMateriaAsignadaPlanificacionSedeVirtualExamen.save()
                                    eMateriaAsignada.visiblehorarioexamen = True

                                    eMateriaAsignada.save()
                                    for mat in MateriaAsignada.objects.filter(status=True, matricula_id=matricula_id):
                                        mat.visiblehorarioexamen = True
                                        mat.save()
                                    if contador >= capacidad:
                                        break
    except Exception as ex:
        print(ex)


def quit_examene_viev():
    from sga.models import MateriaAsignada
    try:
        materiasasig = MateriaAsignada.objects.filter(status=True, matricula__nivel__periodo_id=177,
                                                      matricula__inscripcion__coordinacion_id__in=[1, 2, 3, 4,
                                                                                                   5]).exclude(
            materia__inglesepunemi=True)
        for materia in materiasasig:
            materia.visiblehorarioexamen = False
            materia.save()
            print(materia.matricula.inscripcion, materia.visiblehorarioexamen)
    except Exception as ex:
        print(ex)


def plan_examenes_2023():
    for sede in [10]:
        # if sede == 1:
        #     horas = horas_nivelacion
        # elif sede == 10:
        # horas = horas_nivelacion
        # else:
        #     horas = horas_nivelacion_virtual
        #
        # crear_planificacion_admision_sede_examen(fechasplan, horas, 177, sede)
        # if not sede == 11:
        planificar_pregrado_unemi_v2_regulares(3, sede, 177, 37, fechasplan)
        # else:
        # planificar_pregrado_unemi_virtual_v2(3, sede, 177, 37, fechasplan)
        # planificar_pregrado_unemi_v2_iregulares(3, sede, 177, 37, fechasplan)
    # planificar_transversal()

    # try:
    #     materiasasig = MateriaAsignada.objects.filter(status=True, matricula__nivel__periodo_id=177,
    #                                                   matricula__inscripcion__coordinacion_id__in=[1, 2, 3, 4, 5]).exclude(
    #         materia__inglesepunemi=True)
    #
    #     for materia in materiasasig:
    #         materia.visiblehorarioexamen = True
    #         materia.save()
    #         print(materia.matricula.inscripcion, materia.visiblehorarioexamen)
    #
    # except Exception as ex:
    #     print(ex)


# descrypt('OPPQQRRSSTTUUVVUNRXQ')


def enrolar_new_1():
    for materia in Materia.objects.filter(status=True, nivel__periodo_id=224, idcursomoodle__gt=0,
                                          asignaturamalla__malla__carrera__coordinacion__lte=5):
        print(materia)
        materia.crear_actualizar_estudiantes_curso(moodle, 1)
    print('FIN DEL PROCESO')


# Crea dos hilos para ejecutar las funciones en paralelo
hilo1 = threading.Thread(target=enrolar_new_1)

#
# # Inicia los hilos
# hilo1.start()


# EXAMENEES TRANSVERSALES 2024
fechas_tranversal = [
    (datetime(2024, 1, 6, 0, 0, 0)).date(),
    (datetime(2024, 1, 7, 0, 0, 0)).date(),
]


def planificar_tranversal_realidad_investigacion(limite, sede, periodo_id, detallemodeloevaluativo_id, fechas, grupo):
    limite_x_día = limite
    try:
        materias_grupo = DetalleGrupoAsignatura.objects.filter(status=True, grupo_id=grupo).values('asignatura_id')
        sede_id = sede
        print('PLANIFICANDO SEDE ' + str(sede_id))
        # periodo_id = 153
        # detallemodeloevaluativo_id = 37
        materias_completas = []
        cursor = connections['sga_select'].cursor()
        eMallasIngles = Malla.objects.filter(pk__in=[353, 22]).values_list('id', flat=True)
        matriculasfull = Matricula.objects.filter(materiaasignada__materia__asignaturamalla__transversal=True,
                                                  status=True, materiaasignada__materia__modeloevaluativo_id=27,
                                                  nivel__periodo_id=periodo_id,
                                                  materiaasignada__materia__asignatura_id__in=materias_grupo,
                                                  materiaasignada__materia__asignaturamalla__malla__carrera__modalidad__lte=2)
        eMatriculas = Matricula.objects.filter(materiaasignada__materia__asignaturamalla__transversal=True, status=True,
                                               materiaasignada__materia__modeloevaluativo_id=27,
                                               nivel__periodo_id=periodo_id,
                                               materiaasignada__materia__asignatura_id__in=materias_grupo,
                                               materiaasignada__materia__asignaturamalla__malla__carrera__modalidad__lte=2)
        eMatriculas_exclude_ingles = eMatriculas.annotate(
            total_ingles=Count('materiaasignada__materia__asignaturamalla__id', filter=Q(
                materiaasignada__materia__asignaturamalla__malla_id__in=eMallasIngles.values_list('id', flat=True),
                nivel__periodo_id=periodo_id, status=True)),
            total_general=Count('materiaasignada__materia__asignaturamalla__id',
                                filter=Q(nivel__periodo_id=periodo_id, status=True))).filter(
            total_general=F('total_ingles'))
        ids_exclude = []
        ids_exclude.extend(list(eMatriculas_exclude_ingles.values_list('id', flat=True)))
        eMatriculas = eMatriculas.exclude(pk__in=ids_exclude)
        eMallas = Malla.objects.filter(
            pk__in=eMatriculas.values_list('inscripcion__inscripcionmalla__malla_id', flat=True).distinct()).exclude(
            id__in=eMallasIngles)
        # eCarreras = Carrera.objects.filter(pk__in=eMallas.values_list('carrera__id', flat=True))
        # eCoordinaciones = Coordinacion.objects.filter(pk__in=eCarreras.values_list('coordinacion__id', flat=True)).distinct()
        # for eCoordinacion in eCoordinaciones:
        #     eMallas = eMallas.filter(carrera_id__in=eCoordinacion.carreras().values_list('id', flat=True)).order_by('carrera__nombre', 'inicio')
        eMallas = eMallas.order_by('carrera__nombre', 'inicio')
        for eMalla in eMallas:
            eFechaPlanificacionSedeVirtualExamenes = FechaPlanificacionSedeVirtualExamen.objects.filter(sede_id=sede_id,
                                                                                                        periodo_id=periodo_id,
                                                                                                        fecha__in=fechas).order_by(
                'fecha')
            for eFechaPlanificacionSedeVirtualExamen in eFechaPlanificacionSedeVirtualExamenes:
                fecha = eFechaPlanificacionSedeVirtualExamen.fecha
                eTurnoPlanificacionSedeVirtualExamenes = TurnoPlanificacionSedeVirtualExamen.objects.filter(
                    fechaplanificacion=eFechaPlanificacionSedeVirtualExamen).order_by('horainicio')
                for eTurnoPlanificacionSedeVirtualExamen in eTurnoPlanificacionSedeVirtualExamenes:
                    horainicio = eTurnoPlanificacionSedeVirtualExamen.horainicio
                    horafin = eTurnoPlanificacionSedeVirtualExamen.horafin

                    # eAulaPlanificacionSedeVirtualExamenes = AulaPlanificacionSedeVirtualExamen.objects.filter(turnoplanificacion__fechaplanificacion__sede_id=sede_id,
                    #                                                                                           turnoplanificacion__fechaplanificacion__periodo_id=periodo_id)
                    eAulaPlanificacionSedeVirtualExamenes = AulaPlanificacionSedeVirtualExamen.objects.filter(
                        turnoplanificacion=eTurnoPlanificacionSedeVirtualExamen).exclude(id__lte=29584)
                    eAulaPlanificacionSedeVirtualExamenes_exclude_llenos = eAulaPlanificacionSedeVirtualExamenes.annotate(
                        total_general=Count('materiaasignadaplanificacionsedevirtualexamen__id', filter=Q(
                            materiaasignadaplanificacionsedevirtualexamen__materiaasignada__matricula__nivel__periodo_id=periodo_id,
                            status=True))).filter(total_general=F('aula__capacidad'))
                    eAulaPlanificacionSedeVirtualExamenes.exclude(
                        pk__in=eAulaPlanificacionSedeVirtualExamenes_exclude_llenos.values_list("id", flat=True))
                    eAulaPlanificacionSedeVirtualExamenes = eAulaPlanificacionSedeVirtualExamenes.order_by(
                        'turnoplanificacion__fechaplanificacion__fecha',
                        'turnoplanificacion__horainicio').distinct()
                    # totalAulaSinLlenar = len(eAulaPlanificacionSedeVirtualExamenes.values("id"))
                    contadorAulaSinLlenar = 0
                    # banderaBreakAula = False
                    for eAulaPlanificacionSedeVirtualExamen in eAulaPlanificacionSedeVirtualExamenes:
                        eAula = eAulaPlanificacionSedeVirtualExamen.aula
                        capacidad = eAula.capacidad
                        cantidadad_planificadas = eAulaPlanificacionSedeVirtualExamen.cantidadad_planificadas()
                        eTurnoPlanificacionSedeVirtualExamen = eAulaPlanificacionSedeVirtualExamen.turnoplanificacion
                        # horainicio = eTurnoPlanificacionSedeVirtualExamen.horainicio
                        # horafin = eTurnoPlanificacionSedeVirtualExamen.horafin
                        eFechaPlanificacionSedeVirtualExamen = eTurnoPlanificacionSedeVirtualExamen.fechaplanificacion
                        # fecha = eFechaPlanificacionSedeVirtualExamen.fecha
                        if cantidadad_planificadas < capacidad:
                            print(
                                f"Se procede a planificar en la fecha {fecha} en el horario {horainicio} a {horafin} en el aula {eAula.nombre}")
                            eMateriaAsignadaPlanificacionSedeVirtualExamenes = MateriaAsignadaPlanificacionSedeVirtualExamen.objects.filter(
                                status=True,
                                aulaplanificacion__turnoplanificacion__fechaplanificacion__periodo_id=periodo_id,
                                detallemodeloevaluativo_id=detallemodeloevaluativo_id).values_list(
                                "materiaasignada__id",
                                flat=True)
                            eMateriaAsignadaPlanificacionSedeVirtualExamenesmatriculas = MateriaAsignadaPlanificacionSedeVirtualExamen.objects.filter(
                                status=True,
                                aulaplanificacion__turnoplanificacion__fechaplanificacion__periodo_id=periodo_id,
                                detallemodeloevaluativo_id=detallemodeloevaluativo_id).values_list(
                                "materiaasignada__matricula__id", flat=True)

                            filter_conflicto = (Q(aulaplanificacion__turnoplanificacion__horainicio__lte=horafin,
                                                  aulaplanificacion__turnoplanificacion__horafin__gte=horafin,
                                                  aulaplanificacion__turnoplanificacion__fechaplanificacion__fecha=fecha) |
                                                Q(aulaplanificacion__turnoplanificacion__horainicio__lte=horainicio,
                                                  aulaplanificacion__turnoplanificacion__horafin__gte=horainicio,
                                                  aulaplanificacion__turnoplanificacion__fechaplanificacion__fecha=fecha))
                            eMatriculas = Matricula.objects.filter(
                                pk__in=matriculasfull.values_list("id", flat=True),
                                status=True,
                                retiradomatricula=False,
                                bloqueomatricula=False,
                                nivel__periodo_id=periodo_id,
                                inscripcion__inscripcionmalla__malla=eMalla).exclude(id__in=materias_completas)
                            # eMatriculas = eMatriculas.filter(pk__in=ids_matricula)
                            eMatriculas_exclude_planificadas = eMatriculas.annotate(
                                total_planificadas=Count('materiaasignada__id', filter=Q(
                                    materiaasignada__id__in=eMateriaAsignadaPlanificacionSedeVirtualExamenes,
                                    status=True),
                                                         nivel__periodo_id=periodo_id, status=True),
                                total_general=Count('materiaasignada__id',
                                                    filter=Q(nivel__periodo_id=periodo_id, status=True),
                                                    exclude=Q(materiaasignada__materia__asignatura__id=4837) | Q(
                                                        materiaasignada__materia__asignaturamalla__malla_id__in=eMallasIngles.values_list(
                                                            "id", flat=True)))).filter(
                                Q(total_general=F('total_planificadas')))
                            eMatriculas_exclude_planificadas_x_dia = eMatriculas.annotate(
                                total_planificadas=Count('materiaasignada__id', filter=Q(
                                    materiaasignada__id__in=eMateriaAsignadaPlanificacionSedeVirtualExamenes.filter(
                                        aulaplanificacion__turnoplanificacion__fechaplanificacion__fecha=fecha),
                                    status=True), nivel__periodo_id=periodo_id, status=True)).filter(
                                Q(total_planificadas=limite_x_día))
                            eMatriculas_exclude_ingles = eMatriculas.annotate(
                                total_ingles=Count('materiaasignada__materia__asignaturamalla__id', filter=Q(
                                    materiaasignada__materia__asignaturamalla__malla_id__in=eMallasIngles.values_list(
                                        "id",
                                        flat=True),
                                    nivel__periodo_id=periodo_id, status=True)),
                                total_general=Count('materiaasignada__materia__asignaturamalla__id',
                                                    filter=Q(nivel__periodo_id=periodo_id, status=True))).filter(
                                total_general=F('total_ingles'))
                            ids_exclude = list(
                                eMateriaAsignadaPlanificacionSedeVirtualExamenesmatriculas.filter(filter_conflicto))
                            ids_exclude.extend(list(eMatriculas_exclude_planificadas.values_list('id', flat=True)))
                            ids_exclude.extend(
                                list(eMatriculas_exclude_planificadas_x_dia.values_list('id', flat=True)))
                            ids_exclude.extend(list(eMatriculas_exclude_ingles.values_list('id', flat=True)))
                            eMatriculas = eMatriculas.order_by('inscripcion__inscripcionnivel__nivel__orden',
                                                               'inscripcion__persona__apellido1',
                                                               'inscripcion__persona__apellido2',
                                                               'inscripcion__persona__nombres').distinct()
                            contador = cantidadad_planificadas
                            if not eMatriculas.values("id").exists():
                                contadorAulaSinLlenar += 1

                            if contadorAulaSinLlenar > 0:
                                # banderaBreakAula = True
                                break
                            for eMatricula in eMatriculas:
                                matricula_id = eMatricula.id
                                delet = MateriaAsignadaPlanificacionSedeVirtualExamen.objects.filter(status=True,
                                                                                                     materiaasignada__matricula_id=matricula_id).exclude(
                                    Q(materiaasignada__materia__asignatura__id=4837) |
                                    Q(materiaasignada__materia__asignaturamalla__malla_id__in=eMallasIngles.values_list(
                                        "id", flat=True))).values('id')
                                tmaterias = MateriaAsignada.objects.filter(status=True, matricula_id=matricula_id,
                                                                           materia__asignatura_id__in=materias_grupo).exclude(
                                    Q(materia__asignatura__id=4837) |
                                    Q(materia__asignaturamalla__malla_id__in=eMallasIngles.values_list("id",
                                                                                                       flat=True))).values(
                                    'id')
                                # if len(tmaterias) < 5 or len(tmaterias) > 6:
                                #     continue
                                if len(delet) == len(tmaterias) and not matricula_id in materias_completas:
                                    materias_completas.append(matricula_id)
                                eMateriaAsignadas = MateriaAsignada.objects.filter(status=True, matricula=eMatricula,
                                                                                   materia__asignatura_id__in=materias_grupo)
                                eMateriaAsignadas = eMateriaAsignadas.exclude(
                                    Q(pk__in=eMateriaAsignadaPlanificacionSedeVirtualExamenes) |
                                    Q(materia__asignatura__id=4837) |
                                    Q(materia__asignaturamalla__malla_id__in=eMallasIngles.values_list("id",
                                                                                                       flat=True)))
                                eMateriaAsignadas = eMateriaAsignadas.order_by(
                                    'materia__asignaturamalla__nivelmalla__orden')
                                if eMateriaAsignadas.values("id").exists():
                                    eMateriaAsignada = eMateriaAsignadas.first()
                                    contador += 1
                                    print(
                                        f"------- ({contador}) Se asignada el estudiante {eMateriaAsignada.matricula.inscripcion.persona} en la asignatura {eMateriaAsignada.materia.asignatura.nombre}")
                                    eMateriaAsignadaPlanificacionSedeVirtualExamen = MateriaAsignadaPlanificacionSedeVirtualExamen(
                                        aulaplanificacion=eAulaPlanificacionSedeVirtualExamen,
                                        materiaasignada=eMateriaAsignada,
                                        detallemodeloevaluativo_id=detallemodeloevaluativo_id)
                                    eMateriaAsignadaPlanificacionSedeVirtualExamen.save()
                                    eMateriaAsignada.visiblehorarioexamen = False
                                    eMateriaAsignada.save()
                                    if contador >= capacidad:
                                        break
    except Exception as ex:
        print(ex)


# 1   Investigación
# 2   Realidad
# 3   Lenguaje
def planificar_transversal():
    for grupo in [1, 2, 3]:
        if grupo == 3:
            planificar_tranversal_realidad_investigacion(5, 1, 224, 123, [(datetime(2024, 1, 6, 0, 0, 0)).date()],
                                                         grupo)
        else:
            planificar_tranversal_realidad_investigacion(5, 1, 224, 123, [(datetime(2024, 1, 7, 0, 0, 0)).date()],
                                                         grupo)


# PARA CREAR LA PLANIFICACION DE TRANSVRESALES
def crear_planificacion_transversal():
    horas_6 = [
        [(datetime(2023, 9, 5, 9, 00, 0)).time(), (datetime(2023, 9, 5, 10, 00, 0)).time()],
        [(datetime(2023, 9, 5, 10, 10, 0)).time(), (datetime(2023, 9, 5, 11, 10, 0)).time()],
        [(datetime(2023, 9, 5, 11, 20, 0)).time(), (datetime(2023, 9, 5, 12, 20, 0)).time()]
    ]
    horas_7 = [
        [(datetime(2023, 9, 5, 9, 00, 0)).time(), (datetime(2023, 9, 5, 10, 00, 0)).time()],
        [(datetime(2023, 9, 5, 10, 10, 0)).time(), (datetime(2023, 9, 5, 11, 10, 0)).time()],
        [(datetime(2023, 9, 5, 11, 20, 0)).time(), (datetime(2023, 9, 5, 12, 20, 0)).time()],
        [(datetime(2023, 9, 5, 12, 30, 0)).time(), (datetime(2023, 9, 5, 13, 30, 0)).time()]
    ]
    sede_id = 1
    for fecha in fechas_tranversal:
        print(f"*** FECHA: {fecha}")
        if fecha.strftime('%Y-%m-%d') == '2024-01-06':
            horaslv = horas_6
        else:
            horaslv = horas_7
        eFechaPlanificacionSedeVirtualExamenes = FechaPlanificacionSedeVirtualExamen.objects.filter(sede_id=sede_id,
                                                                                                    periodo_id=224,
                                                                                                    fecha=fecha)
        if not eFechaPlanificacionSedeVirtualExamenes.values("id").exists():
            eFechaPlanificacionSedeVirtualExamen = FechaPlanificacionSedeVirtualExamen(sede_id=sede_id,
                                                                                       periodo_id=224,
                                                                                       fecha=fecha)
            eFechaPlanificacionSedeVirtualExamen.save()

        eFechaPlanificacionSedeVirtualExamen = eFechaPlanificacionSedeVirtualExamenes.first()
        horas = horaslv
        for hora in horas:
            horainicio = hora[0]
            horafin = hora[1]
            print(f"*** FECHA: {fecha} -> hora: {horainicio} - {horafin}")
            eTurnoPlanificacionSedeVirtualExamenes = TurnoPlanificacionSedeVirtualExamen.objects.filter(
                fechaplanificacion=eFechaPlanificacionSedeVirtualExamen, horainicio=horainicio, horafin=horafin)
            if not eTurnoPlanificacionSedeVirtualExamenes.values("id").exists():
                eTurnoPlanificacionSedeVirtualExamen = TurnoPlanificacionSedeVirtualExamen(
                    fechaplanificacion=eFechaPlanificacionSedeVirtualExamen,
                    horainicio=horainicio,
                    horafin=horafin)
                eTurnoPlanificacionSedeVirtualExamen.save()
            else:
                eTurnoPlanificacionSedeVirtualExamen = eTurnoPlanificacionSedeVirtualExamenes.first()

            for eLaboratorioVirtual in LaboratorioVirtual.objects.filter(sedevirtual_id=sede_id, activo=True):
                print(f"*** FECHA: {fecha} -> hora: {horainicio} - {horafin} -> aula: {eLaboratorioVirtual.nombre}")
                eAulaPlanificacionSedeVirtualExamenes = AulaPlanificacionSedeVirtualExamen.objects.filter(
                    turnoplanificacion=eTurnoPlanificacionSedeVirtualExamen, aula=eLaboratorioVirtual)
                if not eAulaPlanificacionSedeVirtualExamenes.values("id").exists():
                    eAulaPlanificacionSedeVirtualExamen = AulaPlanificacionSedeVirtualExamen(
                        turnoplanificacion=eTurnoPlanificacionSedeVirtualExamen,
                        aula=eLaboratorioVirtual)
                    eAulaPlanificacionSedeVirtualExamen.save()
                # else:
                #     eAulaPlanificacionSedeVirtualExamen = eAulaPlanificacionSedeVirtualExamenes.first()
    print(f"********************************FINALIZA PROCESO PLANIFICAR")
    print("")
    print("")
    print("")
    print("")
    print("")
    print(f"********************************PLANIFICANDO EXAMENES")
    planificar_transversal()


def asignar_sede_materias_tranversales():
    for alumno in MateriaAsignada.objects.filter(materia__status=True, materia__nivel__periodo_id=224,
                                                 materia__modeloevaluativo_id=27,
                                                 materia__asignaturamalla__transversal=True, status=True,
                                                 retiramateria=False,
                                                 matricula__retiradomatricula=False,
                                                 matricula__status=True).order_by(
        'matricula__inscripcion__persona__apellido1', 'matricula__inscripcion__persona__apellido2',
        'matricula__inscripcion__persona__nombres').values_list('matricula_id', flat=True).distinct():
        if not MatriculaSedeExamen.objects.filter(status=True, matricula_id=alumno, sede_id=1,
                                                  detallemodeloevaluativo_id=123).exists():
            matriculasede = MatriculaSedeExamen(matricula_id=alumno, sede_id=1, detallemodeloevaluativo_id=123)
            print(alumno, '.......>', matriculasede)
            matriculasede.save()
    crear_planificacion_transversal()


# asignar_sede_materias_tranversales()


def matricula_2_repetidores_admision(periodoactual, periodoanterior, persona, carrera, nivel):
    print("INICIA PROCESO")
    sede = Sede.objects.get(pk=1)
    niveles = Nivel.objects.filter(periodo=periodoactual, nivellibrecoordinacion__coordinacion=ADMISION_ID)
    sesiones = Sesion.objects.filter(pk__in=Nivel.objects.filter(periodo=periodoactual,
                                                                 nivellibrecoordinacion__coordinacion=ADMISION_ID).distinct().values_list(
        'sesion_id'))
    materias = Materia.objects.filter(nivel__in=niveles, status=True)
    esiones = Sesion.objects.filter(pk__in=Nivel.objects.filter(periodo=periodo,
                                                                nivellibrecoordinacion__coordinacion=ADMISION_ID).distinct().values_list(
        'sesion_id'))
    # if carrera:
    #     materias = materias.filter(asignaturamalla__malla__carrera=carrera)
    # print(materias)
    # MATRICULAS DEL PERIODO MAYO A SEPTIEMBRE 2021 EN PREGRADO
    persona_ids = Matricula.objects.filter(status=True, nivel__periodo_id=224, nivelmalla_id=1).exclude(
        inscripcion__coordinacion_id=9).values('inscripcion__persona_id').distinct()
    # persona_ids_aux = Matricula.objects.filter(status=True, nivel__periodo=periodoactual, materiaasignada__isnull=False).values('inscripcion__persona_id').distinct()
    # personas_ids = persona_ids | persona_ids_aux
    personas_ids = persona_ids
    # MATRICULAS DEL PERIODO JUNIO A SEPTIEMBRE 2021 EN ADMISION
    repetidores = Matricula.objects.filter(status=True, aprobado=False, nivel__periodo=periodoanterior, termino=True,
                                           nivelmalla_id=1, inscripcion__coordinacion_id=9).exclude(
        inscripcion__persona_id__in=personas_ids)
    total = len(repetidores.values("id"))
    print(f"Total: {total}")
    contador = 0
    for repe in repetidores:
        materiasrep = MateriaAsignada.objects.filter(matricula=repe, estado_id=2)
        # num_matriculas = repe.inscripcion.historicorecordacademico_set.values('asignatura_id').filter(asignatura_id__in=materiasrep.values('materia__asignatura_id'), fecha__lt=materiasrep.values('materia__nivel__fin').distinct()).annotate(total=Count('asignatura_id')).filter(total=2)
        recordAcademicos = RecordAcademico.objects.filter(inscripcion=repe.inscripcion,
                                                          asignatura_id__in=materiasrep.values(
                                                              'materia__asignatura_id'), status=True, aprobada=False)
        for re in recordAcademicos:
            if re.historicorecordacademico_set.filter(status=True, aprobada=False).count() > 1:
                try:
                    r = Matricula.objects.filter(inscripcion=repe.inscripcion, nivel_id__in=[1516, 1517]).first()
                    if r:
                        print('REPETIDOR -- {}'.format(repe.inscripcion.persona.cedula))
                        r.delete()
                    continue
                except:
                    continue

        contador += 1
        if repe.mismaterias():
            # if repe.materiaasignada_set.filter(status=True).exclude(materiaasignadaretiro__valida=False).count() == 3:
            inscripcion = repe.inscripcion
            if inscripcion.bloqueomatricula:
                inscripcion.bloqueomatricula = False
                inscripcion.save(usuario_id=persona.usuario.id)
            # perfiles_usuarios = PerfilUsuario.objects.filter(persona=inscripcion.persona, inscripcion=inscripcion)
            # if perfiles_usuarios.exists():
            #     perfil = perfiles_usuarios.first()
            #     perfil.visible = False
            #     perfil.save(usuario_id=persona.usuario.id)
            mimalla = inscripcion.mi_malla()
            mallaacutualcarrera = Malla.objects.filter(status=True, carrera=inscripcion.carrera, validamatricula=True,
                                                       vigente=True).first()
            # if mallaacutualcarrera and mimalla and (not mallaacutualcarrera.id == mimalla.id):
            #     malla = inscripcion.inscripcionmalla_set.filter(status=True)
            #     malla.delete()
            #     im = InscripcionMalla(inscripcion=inscripcion, malla=mallaacutualcarrera)
            #     im.save()
            #     inscripcion.actualizar_creditos()
            #     mimalla = im.malla
            if not Materia.objects.filter(nivel__periodo=periodoactual, asignaturamalla__malla=mimalla,
                                          asignaturamalla__malla__carrera=inscripcion.carrera,
                                          nivel__sesion=inscripcion.sesion).exists():
                return False
            nivel_id = 0
            for sesion in sesiones:
                # MODALIDAD EN LINEA (LE AGREGO LA SESION DE EN LINEA)
                if sesion.id in [13] and inscripcion.modalidad.id in [3]:
                    inscripcion.sesion = sesion
                    nivel_id = 1516
                # MODALIDAD EN SEMIPRESENCIAL O PRESENCIAL (LE AGREGO EL ID DE LA SESION DE FNSEMANA)
                elif sesion.id in [7, 11, 12] and inscripcion.modalidad.id in [1, 2]:
                    inscripcion.sesion = sesion
                    nivel_id = 1517
            nivel = Nivel.objects.get(id=nivel_id)
            matricula = Matricula.objects.filter(inscripcion=inscripcion, nivel_id=nivel).first()

            # if not repe.inscripcion.matricula_periodo(periodoactual):
            if not matricula:
                matricula = Matricula(inscripcion=inscripcion,
                                      nivel=nivel,
                                      pago=False,
                                      iece=False,
                                      becado=False,
                                      porcientobeca=0,
                                      fecha=datetime.now().date(),
                                      hora=datetime.now().time(),
                                      fechatope=fechatope(datetime.now().date()),
                                      automatriculaadmision=True,
                                      fechaautomatriculaadmision=datetime.now(),
                                      termino=False)
                matricula.save(usuario_id=persona.usuario.id)
            # else:
            #     matricula = Matricula.objects.get(inscripcion=inscripcion, nivel=nivel)
            de = Matricula.objects.filter(inscripcion=inscripcion, nivel_id__in=[1516, 1517]).exclude(
                id=matricula.id).first()
            if de:
                de.delete()
            print(f"{total}/{contador} --> {matricula.__str__()}")
            if not MateriaAsignada.objects.values("id").filter(matricula=matricula).exists() or 1 == 1:
                materias = MateriaAsignada.objects.filter(matricula=repe, estado_id=2)
                eRecordAcademicos = RecordAcademico.objects.filter(inscripcion=repe.inscripcion,
                                                                   materiaregular__id__in=materias.values_list(
                                                                       'materia__id', flat=True), status=True)
                asignaturas = RecordAcademico.objects.filter(inscripcion=repe.inscripcion,
                                                             materiaregular__id__in=materias.values_list('materia_id',
                                                                                                         flat=True),
                                                             status=True).values_list('asignatura_id', flat=True)
                materias_actuales = []
                paralelos = []
                for eRecordAcademico in eRecordAcademicos:
                    # print(eRecordAcademico.asignatura.id)
                    for d in asignaturas:
                        # print(d)
                        if eRecordAcademico.asignatura_id == d:
                            asignatura_actual_id = d
                            mats = Materia.objects.filter(nivel=nivel, asignaturamalla__malla=mimalla,
                                                          asignatura_id=asignatura_actual_id)
                            if not MateriaAsignada.objects.filter(materia_id__in=mats.values_list('id', flat=True),
                                                                  matricula=matricula).exists():
                                for mat in mats:
                                    cupo = mat.cupo
                                    matriculados = MateriaAsignada.objects.filter(materia=mat).count()
                                    if matriculados <= cupo:
                                        paralelos.append(mat.paralelomateria.id)
                                        materias_actuales.append(mat.id)
                                        break
                # print(materias_actuales)
                for iMat in materias_actuales:
                    materia = Materia.objects.get(pk=iMat)
                    if not MateriaAsignada.objects.filter(matricula=matricula, materia=materia).exists():
                        num_matriculas = repe.inscripcion.historicorecordacademico_set.values('id').filter(
                            asignatura=materia.asignatura, fecha__lt=materia.nivel.fin).count() + 1
                        materiaasignada = MateriaAsignada(matricula=matricula,
                                                          materia=materia,
                                                          notafinal=0,
                                                          asistenciafinal=0,
                                                          cerrado=False,
                                                          matriculas=num_matriculas,
                                                          observaciones='',
                                                          estado_id=NOTA_ESTADO_EN_CURSO,
                                                          cobroperdidagratuidad=True)
                        materiaasignada.save(usuario_id=persona.usuario.id)
                        materiaasignada.asistencias()
                        materiaasignada.evaluacion()
                        materiaasignada.mis_planificaciones()
                        materiaasignada.save(usuario_id=persona.usuario.id)
                        # print(materiaasignada)
            matricula.actualizar_horas_creditos()
            matricula.estado_matricula = 2
            matricula.save(usuario_id=persona.usuario.id)
            matricula.calcula_nivel()
            inscripcion.actualizar_nivel()
            # if not materia.asignatura.id in [4837]:

            if inscripcion.sesion_id == 13:
                tiporubromatricula = TipoOtroRubro.objects.get(pk=3019)
            else:
                tiporubromatricula = TipoOtroRubro.objects.get(pk=3011)

            # num_materias = MateriaAsignada.objects.filter(matricula=matricula, cobroperdidagratuidad=True).exclude(materia__asignatura_id=4837).count()
            num_materias = MateriaAsignada.objects.filter(matricula=matricula, cobroperdidagratuidad=True).count()
            if num_materias > 0:
                valor_x_materia = 20
                valor_total = num_materias * valor_x_materia
                matricula.estado_matricula = 1
                matricula.save(usuario_id=persona.usuario.id)
                rubro1 = Rubro.objects.filter(persona=inscripcion.persona, matricula=matricula).first()
                if not rubro1:
                    rubro1 = Rubro(tipo=tiporubromatricula,
                                   persona=inscripcion.persona,
                                   matricula=matricula,
                                   nombre=tiporubromatricula.nombre + ' - ' + periodoactual.nombre,
                                   cuota=1,
                                   fecha=datetime.now().date(),
                                   # fechavence=datetime.now().date() + timedelta(days=22),
                                   fechavence=(datetime(2023, 11, 5, 0, 0, 0)).date(),
                                   valor=valor_total,
                                   iva_id=1,
                                   valoriva=0,
                                   valortotal=valor_total,
                                   saldo=valor_total,
                                   cancelado=False)
                else:
                    if not rubro1.cancelado:
                        rubro1.valor = valor_total
                        rubro1.valortotal = valor_total
                        rubro1.saldo = valor_total
                rubro1.save(usuario_id=persona.usuario.id)
                # print(rubro1)
    return


actual = Periodo.objects.get(id=224)
anterior = Periodo.objects.get(id=177)
persona = Persona.objects.get(id=1)
carrera = Carrera.objects.get(id=104)
nivel = Nivel.objects.get(id=1516)


# matricula_2_repetidores_admision(actual, anterior, persona, carrera, nivel)


def borrar_con_materias_antiguas(periodoactual, periodoanterior, persona):
    try:
        print("INICIA PROCESO")
        sede = Sede.objects.get(pk=1)
        niveles = Nivel.objects.filter(periodo=periodoactual, nivellibrecoordinacion__coordinacion=ADMISION_ID)
        sesiones = Sesion.objects.filter(pk__in=Nivel.objects.filter(periodo=periodoactual,
                                                                     nivellibrecoordinacion__coordinacion=ADMISION_ID).distinct().values_list(
            'sesion_id'))
        materias = Materia.objects.filter(nivel__in=niveles, status=True)
        esiones = Sesion.objects.filter(pk__in=Nivel.objects.filter(periodo=periodo,
                                                                    nivellibrecoordinacion__coordinacion=ADMISION_ID).distinct().values_list(
            'sesion_id'))
        persona_ids = Matricula.objects.filter(status=True, nivel__periodo_id=224, nivelmalla_id=1).exclude(
            inscripcion__coordinacion_id=9).values('inscripcion__persona_id').distinct()
        personas_ids = persona_ids
        # MATRICULAS DEL PERIODO JUNIO A SEPTIEMBRE 2021 EN ADMISION
        repetidores = Matricula.objects.filter(status=True, aprobado=False, nivel__periodo=periodoanterior,
                                               termino=True, nivelmalla_id=1, inscripcion__coordinacion_id=9).exclude(
            inscripcion__persona_id__in=personas_ids)
        total = len(repetidores.values("id"))
        print(f"Total: {total}")
        contador = 0
        for repe in repetidores:
            contador += 1
            materiasrep = MateriaAsignada.objects.filter(matricula=repe, estado_id=2)
            # num_matriculas = repe.inscripcion.historicorecordacademico_set.values('asignatura_id').filter(asignatura_id__in=materiasrep.values('materia__asignatura_id'), fecha__lt=materiasrep.values('materia__nivel__fin').distinct()).annotate(total=Count('asignatura_id')).filter(total=2)
            recordAcademicos = RecordAcademico.objects.filter(inscripcion=repe.inscripcion,
                                                              asignatura_id__in=materiasrep.values(
                                                                  'materia__asignatura_id'), status=True,
                                                              aprobada=False)
            for re in recordAcademicos:
                if re.historicorecordacademico_set.filter(status=True, aprobada=False).count() > 1:
                    try:
                        r = Matricula.objects.filter(inscripcion=repe.inscripcion, nivel_id__in=[1516, 1517]).first()
                        if r:
                            # print('REPETIDOR -- {}'.format(repe.inscripcion.persona.cedula))
                            print(f"{total}/{contador} --> **** ELIMINANDO MATRICULA ********** {r.__str__()}")
                            r.delete()
                        break
                    except:
                        break
                else:
                    print(f"{total}/{contador} --> NO APLICA ELIMINACION --> {repe.__str__()}")
                    break
            if not recordAcademicos:
                print(f"{total}/{contador} --> NO APLICA ELIMINACION --> {repe.__str__()}")
            # print(f"{total}/{contador} --> {r.__str__()}")


    except Exception as ex:
        print(ex)


# borrar_con_materias_antiguas(actual, anterior, persona)

def asignar_rubros():
    try:
        tienenpagos = []
        persona = Persona.objects.get(id=1)
        periodoactual = Periodo.objects.get(id=224)
        valor_x_materia = 20
        c = 0
        matriculas = Matricula.objects.filter(cuposenescyt=False, status=True, nivel_id__in=[1516, 1517]).order_by('-inscripcion_id','inscripcion__estado_gratuidad')
        print('INICIANDO PROCESO')
        print(f'TOTAL DE REGISTROS {len(matriculas)}')
        for matricula in matriculas:
            inscripcion = matricula.inscripcion
            num_materias = MateriaAsignada.objects.filter(matricula=matricula).count()
            c += 1
            print(f"{c} -->: {inscripcion}")
            if inscripcion.sesion_id == 13:
                tiporubromatricula = TipoOtroRubro.objects.get(pk=3019)
            else:
                tiporubromatricula = TipoOtroRubro.objects.get(pk=3011)
            if num_materias > 0:
                valor_total = num_materias * valor_x_materia
                # matricula.estado_matricula = 1
                # matricula.save(usuario_id=persona.usuario.id)
                rubro1 = Rubro.objects.filter(persona=inscripcion.persona, matricula=matricula).first()
                repetidor = Matricula.objects.filter(status=True, inscripcion=inscripcion, aprobado=False,
                                                     nivel__periodo_id=177, termino=True, nivelmalla_id=1,
                                                     inscripcion__coordinacion_id=9).first()
                if not rubro1:
                    if inscripcion.estado_gratuidad == 3 or repetidor:
                        print(f"Creando Rurbo")
                        rubro1 = Rubro(tipo=tiporubromatricula,
                                       persona=inscripcion.persona,
                                       matricula=matricula,
                                       nombre=tiporubromatricula.nombre + ' - ' + periodoactual.nombre,
                                       cuota=1,
                                       fecha=datetime.now().date(),
                                       # fechavence=datetime.now().date() + timedelta(days=22),
                                       fechavence=datetime.now().date() + timedelta(days=34),
                                       valor=valor_total,
                                       iva_id=1,
                                       valoriva=0,
                                       valortotal=valor_total,
                                       saldo=valor_total,
                                       cancelado=False)
                        rubro1.save(usuario_id=persona.usuario.id)
                        matricula.estado_matricula = 3
                        matricula.save(usuario_id=persona.usuario.id)
                    # print(rubro1)
                else:
                    if inscripcion.estado_gratuidad == 1 and rubro1:
                        if not rubro1.tiene_pagos() and not repetidor:
                            print(f"ELIMINANDO RUBRO, ALUMNO TIENE GRATUIDAD")
                            rubro1.delete()
                    if inscripcion.estado_gratuidad == 1 and rubro1:
                        if rubro1.tiene_pagos() and not repetidor:
                            tienenpagos.append(inscripcion.id)
                    if not rubro1.valortotal == valor_total:
                        if inscripcion.estado_gratuidad == 3 or repetidor:
                            print(f"ACTUALIZANDO RUBRO, VALOR INCORRECTO")
                            rubro1.valor = float(valor_total)
                            rubro1.valortotal = float(valor_total)
                            rubro1.saldo = float(valor_total - rubro1.total_pagado())
                            rubro1.fechavence = datetime.now().date() + timedelta(days=34)
                            rubro1.save(usuario_id=persona.usuario.id)
                # if repetidor:
                #     materiasrep = MateriaAsignada.objects.filter(matricula=repetidor, estado_id=2)
                #     # num_matriculas = repe.inscripcion.historicorecordacademico_set.values('asignatura_id').filter(asignatura_id__in=materiasrep.values('materia__asignatura_id'), fecha__lt=materiasrep.values('materia__nivel__fin').distinct()).annotate(total=Count('asignatura_id')).filter(total=2)
                #     recordAcademicos = RecordAcademico.objects.filter(inscripcion=inscripcion,
                #                                                       asignatura_id__in=materiasrep.values(
                #                                                           'materia__asignatura_id'), status=True,
                #                                                       aprobada=False)
                #     for re in recordAcademicos:
                #         if re.historicorecordacademico_set.filter(status=True, aprobada=False).count() > 1:
                #             try:
                #                 if not rubro1.tiene_pagos():
                #                     # print('REPETIDOR -- {}'.format(repe.inscripcion.persona.cedula))
                #                     print(f"**** ELIMINANDO MATRICULA ********** {matricula.__str__()}")
                #                     matricula.delete()
                #                 break
                #             except:
                #                 break

        print(tienenpagos)
    except Exception as ex:
        print(ex)

#asignar_rubros()

def matriculas_admison():
    try:
        periodoactual = Periodo.objects.get(id=224)
        persona = Persona.objects.get(id=1)
        sesiones = Sesion.objects.filter(pk__in=Nivel.objects.filter(periodo=periodoactual,nivellibrecoordinacion__coordinacion=ADMISION_ID).distinct().values_list('sesion_id'))

        for matr in Matricula.objects.filter(status=True, nivel_id__in=[1516,1517]):
        #for matr in Matricula.objects.filter(pk=792683):
            eInscripcion = matr.inscripcion
            if eInscripcion:
                perfiles_usuarios = PerfilUsuario.objects.filter(persona=eInscripcion.persona, inscripcion=eInscripcion)
                if not perfiles_usuarios.values("id").exists():
                    eInscripcion.persona.crear_perfil(inscripcion=eInscripcion, visible=False)

                nivel_id = 0
                for sesion in sesiones:
                    # MODALIDAD EN LINEA (LE AGREGO LA SESION DE EN LINEA)
                    if sesion.id in [13] and eInscripcion.modalidad.id in [3]:
                        eInscripcion.sesion = sesion
                        nivel_id = 1516
                    # MODALIDAD EN SEMIPRESENCIAL O PRESENCIAL (LE AGREGO EL ID DE LA SESION DE FNSEMANA)
                    elif sesion.id in [7, 11, 12] and eInscripcion.modalidad.id in [1, 2]:
                        eInscripcion.sesion = sesion
                        nivel_id = 1517

                eInscripcion.save(usuario_id=persona.usuario.id)
                # hago la matriculación
                mimalla = eInscripcion.malla_inscripcion()
                mallaacutualcarrera = Malla.objects.filter(status=True, carrera=eInscripcion.carrera, validamatricula=True,
                                                           vigente=True).first()
                if mallaacutualcarrera and mimalla and (not mallaacutualcarrera.id == mimalla.malla_id):
                    malla = eInscripcion.inscripcionmalla_set.filter(status=True)
                    malla.delete()
                    im = InscripcionMalla(inscripcion=eInscripcion, malla=mallaacutualcarrera)
                    im.save()
                    eInscripcion.actualizar_creditos()
                    mimalla = im
                print(nivel_id)
                # nivel = Nivel.objects.get(periodo=periodo, sesion=eInscripcion.sesion, sede=sede)
                nivel = Nivel.objects.get(pk=nivel_id)

                if Materia.objects.filter(nivel__periodo=periodoactual, asignaturamalla__malla=mimalla.malla,
                                          asignaturamalla__malla__carrera=eInscripcion.carrera,
                                          nivel__sesion=eInscripcion.sesion).exists():
                    # if not eInscripcion.matricula_periodo(periodo):
                    matricula = Matricula.objects.filter(inscripcion=eInscripcion, nivel=nivel).first()
                    # if not matricula:
                    #     matricula = Matricula(inscripcion=eInscripcion,
                    #                           nivel=nivel,
                    #                           pago=False,
                    #                           iece=False,
                    #                           becado=False,
                    #                           porcientobeca=0,
                    #                           fecha=datetime.now().date(),
                    #                           hora=datetime.now().time(),
                    #                           fechatope=fechatope(datetime.now().date()),
                    #                           automatriculaadmision=True,
                    #                           fechaautomatriculaadmision=datetime.now())
                    #     matricula.save(usuario_id=persona.usuario.id)
                    # else:
                    #     matricula = Matricula.objects.get(inscripcion=eInscripcion, nivel=nivel)
                    print(matricula)
                    eMateriaAsignadas = MateriaAsignada.objects.filter(matricula=matricula)
                    if not eMateriaAsignadas.values("id").exists() or eMateriaAsignadas.values("id").count() < 3:
                        # paralelos = Materia.objects.filter(nivel__periodo=periodo, asignaturamalla__malla=mimalla.malla, asignaturamalla__malla__carrera=inscripcion.carrera, nivel__sesion=inscripcion.sesion).values_list('paralelo').distinct().order_by('paralelo')
                        paralelos = Materia.objects.filter(nivel__periodo=periodoactual, asignaturamalla__malla=mimalla.malla,
                                                           asignaturamalla__malla__carrera=eInscripcion.carrera,
                                                           nivel__sesion=eInscripcion.sesion).values_list(
                            'paralelomateria').distinct()

                        if paralelos.values("id").exists():
                            paralelo_atomar = None
                            tiene_cupo_paralelo = False
                            for paralelo in paralelos:
                                tiene_cupo_paralelo_aux = True
                                for mat in Materia.objects.filter(nivel__periodo=periodoactual, paralelomateria=paralelo,
                                                                  asignaturamalla__malla=mimalla.malla,
                                                                  asignaturamalla__malla__carrera=eInscripcion.carrera,
                                                                  nivel__sesion=eInscripcion.sesion):
                                    if MateriaAsignada.objects.filter(materia=mat).count() + 1 > mat.cupo:
                                        tiene_cupo_paralelo_aux = False
                                        break
                                if tiene_cupo_paralelo_aux:
                                    paralelo_atomar = paralelo
                                    tiene_cupo_paralelo = True
                                    break
                            if tiene_cupo_paralelo:
                                materias_c = Materia.objects.filter(nivel__periodo=periodoactual, paralelomateria=paralelo_atomar,
                                                                    asignaturamalla__malla=mimalla.malla,
                                                                    asignaturamalla__malla__carrera=eInscripcion.carrera,
                                                                    nivel__sesion=eInscripcion.sesion)
                                for materia in materias_c:
                                    if not MateriaAsignada.objects.values('id').filter(matricula=matricula,
                                                                                       materia=materia).exists():
                                        matriculas = matricula.inscripcion.historicorecordacademico_set.values('id').filter(
                                            asignatura=materia.asignatura, fecha__lt=materia.nivel.fin).count() + 1
                                        materiaasignada = MateriaAsignada(matricula=matricula,
                                                                          materia=materia,
                                                                          notafinal=0,
                                                                          asistenciafinal=0,
                                                                          cerrado=False,
                                                                          matriculas=matriculas,
                                                                          observaciones='',
                                                                          estado_id=NOTA_ESTADO_EN_CURSO,
                                                                          cobroperdidagratuidad=eInscripcion.gratuidad)
                                        materiaasignada.save(usuario_id=persona.usuario.id)
                                        materiaasignada.asistencias()
                                        materiaasignada.evaluacion()
                                        materiaasignada.mis_planificaciones()
                                        materiaasignada.save(usuario_id=persona.usuario.id)
                                        print(materiaasignada)

                        matricula.actualizar_horas_creditos()
                        matricula.estado_matricula = 2
                        matricula.save(usuario_id=persona.usuario.id)
                        matricula.calcula_nivel()
                        eInscripcion.actualizar_nivel()
                        if eInscripcion.estado_gratuidad == 3:
                            if eInscripcion.sesion_id == 13:
                                tiporubromatricula = TipoOtroRubro.objects.get(pk=3019)
                            else:
                                tiporubromatricula = TipoOtroRubro.objects.get(pk=3011)
                            eMateriaAsignadas.update(cobroperdidagratuidad=True)
                            if matricula.tipomatricula_id == 1:
                                matricula.estado_matricula = 2
                                matricula.save(usuario_id=persona.usuario.id)

                            num_materias = MateriaAsignada.objects.filter(matricula=matricula,
                                                                          cobroperdidagratuidad=True).count()
                            valor_x_materia = 20
                            valor_total = num_materias * valor_x_materia

                            if not Rubro.objects.filter(persona=eInscripcion.persona, matricula=matricula).exists():
                                print(eInscripcion.gratuidad)
                                rubro1 = Rubro(tipo=tiporubromatricula,
                                               persona=eInscripcion.persona,
                                               matricula=matricula,
                                               nombre=tiporubromatricula.nombre + ' - ' + periodoactual.nombre,
                                               cuota=1,
                                               fecha=datetime.now().date(),
                                               # fechavence=datetime.now().date() + timedelta(days=22),
                                               fechavence = datetime.now().date(),
                                               valor=valor_total,
                                               iva_id=1,
                                               valoriva=0,
                                               valortotal=valor_total,
                                               saldo=valor_total,
                                               cancelado=False)
                                rubro1.save(usuario_id=persona.usuario.id)
                                print(rubro1)
                            else:
                                rubro1 = Rubro.objects.filter(persona=eInscripcion.persona, matricula=matricula)[0]
                                rubro1.tipo = tiporubromatricula
                                rubro1.nombre = tiporubromatricula.nombre + ' - ' + periodoactual.nombre
                                rubro1.cuota = 1
                                rubro1.fecha = datetime.now().date()
                                # rubro1.fechavence = datetime.now().date() + timedelta(days=22)
                                rubro1.fechavence = datetime.now().date()
                                rubro1.valor = valor_total
                                rubro1.iva_id = 1
                                rubro1.valoriva = 0
                                rubro1.valortotal = valor_total
                                rubro1.saldo = valor_total
                                rubro1.cancelado = False
                                rubro1.save(usuario_id=persona.usuario.id)
                                print(rubro1)
    except Exception as ex:
        print(ex)
        import sys
        print('Error on line {} - {}'.format(sys.exc_info()[-1].tb_lineno, ex))

# matriculas_admison()


def eliminar_materias_asignadas():
    materiasasignadas= MateriaAsignada.objects.filter(status=True, matricula__nivel__periodo__id=224, matricula__nivel__id__in=[1516,1517], materia__nivel__id__in=[1481,1482])
    for materia in materiasasignadas:
        matricula=materia.matricula
        rubros = Rubro.objects.filter(status=True, matricula=matricula)
        for rubro in rubros:
            if not rubro.tiene_pagos():
                rubro.delete()
        materia.delete()

#eliminar_materias_asignadas()



def notas_ingles():
    print('Inicio del proceso')
    miarchivo = openpyxl.load_workbook("ingles_2023.xlsx")
    lista = miarchivo.get_sheet_by_name('notas')
    totallista = lista.rows
    a = 0
    for filas in totallista:
        a += 1
        if a > 1:
            cedula = filas[0].value
            nota = filas[1].value
            carrera = filas[2].value
            fecha = filas[3].value
            inscripcion = Inscripcion.objects.filter(status=True, carrera_id=carrera, persona__cedula=cedula).first()
            for list in GrupoInscripcion.objects.filter(status=True, grupo_id__gte=5, inscripcion__persona__cedula=cedula):#mayor a 5
                record = RecordAcademico.objects.filter(status=True, grupoinscripcion=list,
                                                        observaciones__icontains='proceso RE-CEP-EPUNEMI-2023-03')
                if inscripcion:
                    if not list.inscripcion == inscripcion:
                        print('CAMBIANDO DE CARRERA')
                        if record:
                            for re in record:
                                re.delete()
                        if inscripcion:
                            list.inscripcion = inscripcion
                            list.save()
                            print(list.inscripcion)
                            malla = list.inscripcion.inscripcionmalla_set.values_list('malla_id', flat=True).filter(status=True)
                            modulomalla = ModuloMalla.objects.filter(status=True, malla_id=malla[0], orden__in=[6, 7])
                            if not modulomalla:
                                print('NO TIENE MODULOS EN LA MALLA')
                            for asig in modulomalla:
                                print(asig.asignatura)
                                if not RecordAcademico.objects.filter(inscripcion=list.inscripcion,
                                                                      asignatura=asig.asignatura).exists():
                                    record = RecordAcademico(inscripcion=list.inscripcion,
                                                             asignatura=asig.asignatura,
                                                             grupoinscripcion=list,
                                                             modulomalla=list.inscripcion.asignatura_en_modulomalla(asig.asignatura),
                                                             nota=nota,
                                                             asistencia=100,
                                                             fecha=fecha,
                                                             aprobada=True,
                                                             noaplica=False,
                                                             convalidacion=False,
                                                             pendiente=False,
                                                             creditos=asig.creditos,
                                                             horas=asig.horas,
                                                             homologada=False,
                                                             valida=True,
                                                             validapromedio=False,
                                                             observaciones='proceso RE-CEP-EPUNEMI-2023-03',
                                                             suficiencia=True)
                                    record.save()
                                else:
                                    record = RecordAcademico.objects.filter(inscripcion=list.inscripcion,
                                                                         asignatura=asig.asignatura).first()
                                if not record.historicorecordacademico_set.filter(status=True).exists():
                                    nuevohistorico = HistoricoRecordAcademico(recordacademico=record,
                                                                              inscripcion=record.inscripcion,
                                                                              modulomalla=record.modulomalla,
                                                                              asignaturamalla=record.asignaturamalla,
                                                                              asignatura=record.asignatura if record.asignatura else record.asignaturamallahistorico.asignatura,
                                                                              grupoinscripcion=record.grupoinscripcion,
                                                                              nota=nota,
                                                                              asistencia=record.asistencia,
                                                                              sinasistencia=record.sinasistencia,
                                                                              fecha=record.fecha,
                                                                              noaplica=record.noaplica,
                                                                              aprobada=record.aprobada,
                                                                              convalidacion=record.convalidacion,
                                                                              homologada=record.homologada,
                                                                              pendiente=record.pendiente,
                                                                              creditos=record.creditos,
                                                                              horas=record.horas,
                                                                              valida=record.valida,
                                                                              validapromedio=record.validapromedio,
                                                                              materiaregular=record.materiaregular,
                                                                              materiacurso=record.materiacurso,
                                                                              observaciones=record.observaciones,
                                                                              completonota=record.completonota,
                                                                              completoasistencia=record.completoasistencia,
                                                                              suficiencia=record.suficiencia)
                                    nuevohistorico.save()
                    else:
                        if not record:
                            if inscripcion:
                                list.inscripcion = inscripcion
                                list.save()
                                print(list.inscripcion)
                                malla = list.inscripcion.inscripcionmalla_set.values_list('malla_id', flat=True).filter(
                                    status=True)
                                modulomalla = ModuloMalla.objects.filter(status=True, malla_id=malla[0],
                                                                         orden__in=[6, 7])
                                if not modulomalla:
                                    print('NO TIENE MODULOS EN LA MALLA')
                                for asig in modulomalla:
                                    print(asig.asignatura)
                                    if not RecordAcademico.objects.filter(inscripcion=list.inscripcion,
                                                                          asignatura=asig.asignatura).exists():
                                        record = RecordAcademico(inscripcion=list.inscripcion,
                                                                 asignatura=asig.asignatura,
                                                                 grupoinscripcion=list,
                                                                 modulomalla=list.inscripcion.asignatura_en_modulomalla(
                                                                     asig.asignatura),
                                                                 nota=nota,
                                                                 asistencia=100,
                                                                 fecha=fecha,
                                                                 aprobada=True,
                                                                 noaplica=False,
                                                                 convalidacion=False,
                                                                 pendiente=False,
                                                                 creditos=asig.creditos,
                                                                 horas=asig.horas,
                                                                 homologada=False,
                                                                 valida=True,
                                                                 validapromedio=False,
                                                                 observaciones='proceso RE-CEP-EPUNEMI-2023-03',
                                                                 suficiencia=True)
                                        record.save()
                                    else:
                                        record = RecordAcademico.objects.filter(inscripcion=list.inscripcion,
                                                                                asignatura=asig.asignatura).first()
                                    if not record.historicorecordacademico_set.filter(status=True).exists():
                                        nuevohistorico = HistoricoRecordAcademico(recordacademico=record,
                                                                                  inscripcion=record.inscripcion,
                                                                                  modulomalla=record.modulomalla,
                                                                                  asignaturamalla=record.asignaturamalla,
                                                                                  asignatura=record.asignatura if record.asignatura else record.asignaturamallahistorico.asignatura,
                                                                                  grupoinscripcion=record.grupoinscripcion,
                                                                                  nota=nota,
                                                                                  asistencia=record.asistencia,
                                                                                  sinasistencia=record.sinasistencia,
                                                                                  fecha=record.fecha,
                                                                                  noaplica=record.noaplica,
                                                                                  aprobada=record.aprobada,
                                                                                  convalidacion=record.convalidacion,
                                                                                  homologada=record.homologada,
                                                                                  pendiente=record.pendiente,
                                                                                  creditos=record.creditos,
                                                                                  horas=record.horas,
                                                                                  valida=record.valida,
                                                                                  validapromedio=record.validapromedio,
                                                                                  materiaregular=record.materiaregular,
                                                                                  materiacurso=record.materiacurso,
                                                                                  observaciones=record.observaciones,
                                                                                  completonota=record.completonota,
                                                                                  completoasistencia=record.completoasistencia,
                                                                                  suficiencia=record.suficiencia)
                                        nuevohistorico.save()
                        else:
                            print('ACTUALIZANDO FECHA DE EXÁMEN', list.inscripcion)
                            record.update(fecha=fecha)
                            for re in record:
                                historico = re.historicorecordacademico_set.filter(status=True)
                                historico.update(fecha=fecha)



# notas_ingles()


fechas_examenes_2023_milagro = [
    (datetime(2024, 1, 23, 0, 0, 0)).date(),
    (datetime(2024, 1, 24, 0, 0, 0)).date()
]
def planificar_pregrado_unemi_2023(limite_x_día, sede, periodo_id, detallemodeloevaluativo_id, fechas):
        try:
            sede_id = sede
            print('PLANIFICANDO SEDE ' + str(sede_id))
            # periodo_id = 153
            # detallemodeloevaluativo_id = 37
            materias_completas = []
            cursor = connections['sga_select'].cursor()
            eMallasIngles = Malla.objects.filter(pk__in=[353, 22]).values_list('id', flat=True)
            eMallasMedicina = Malla.objects.filter(carrera__id__in=[223]).values_list('id', flat=True)
            eMatriculaSedeExamenes = MatriculaSedeExamen.objects.filter(
                detallemodeloevaluativo_id=detallemodeloevaluativo_id,
                status=True, matricula__status=True,
                matricula__retiradomatricula=False,
                matricula__nivel__periodo_id=periodo_id,
                matricula__inscripcion__carrera__modalidad=3).filter(
                matricula__materiaasignada__materia__asignaturamalla__nivelmalla_id=9)
            # if DEBUG:
            # MateriaAsignadaPlanificacionSedeVirtualExamen.objects.filter(status=True,
            #                                                              aulaplanificacion__turnoplanificacion__fechaplanificacion__periodo_id=periodo_id,
            #                                                              aulaplanificacion__turnoplanificacion__fechaplanificacion__sede_id=sede_id,
            #                                                              detallemodeloevaluativo_id=detallemodeloevaluativo_id).delete()
            # eMatriculaSedeExamenes.update(sede_id=sede_id)

            eMatriculas = Matricula.objects.filter(
                pk__in=eMatriculaSedeExamenes.values_list("matricula__id", flat=True),
                status=True, retiradomatricula=False,
                nivel__periodo_id=periodo_id).filter(materiaasignada__materia__asignaturamalla__nivelmalla_id=9)
            eMatriculas_exclude_ingles = eMatriculas.annotate(
                total_ingles=Count('materiaasignada__materia__asignaturamalla__id', filter=Q(
                    materiaasignada__materia__asignaturamalla__malla_id__in=eMallasIngles.values_list('id', flat=True),
                    nivel__periodo_id=periodo_id, status=True)),
                total_general=Count('materiaasignada__materia__asignaturamalla__id',
                                    filter=Q(nivel__periodo_id=periodo_id, status=True))).filter(
                total_general=F('total_ingles'))
            ids_exclude = []
            ids_exclude.extend(list(eMatriculas_exclude_ingles.values_list('id', flat=True)))
            eMatriculas = eMatriculas.exclude(pk__in=ids_exclude)
            eMallas = Malla.objects.filter(carrera__modalidad=3,
                                           pk__in=eMatriculas.values_list('inscripcion__inscripcionmalla__malla_id',
                                                                          flat=True).distinct()).exclude(
                id__in=eMallasIngles)
            # eCarreras = Carrera.objects.filter(pk__in=eMallas.values_list('carrera__id', flat=True))
            # eCoordinaciones = Coordinacion.objects.filter(pk__in=eCarreras.values_list('coordinacion__id', flat=True)).distinct()
            # for eCoordinacion in eCoordinaciones:
            #     eMallas = eMallas.filter(carrera_id__in=eCoordinacion.carreras().values_list('id', flat=True)).order_by('carrera__nombre', 'inicio')
            eMallas = eMallas.order_by('carrera__nombre', 'inicio')
            for eMalla in eMallas:
                print(eMalla.carrera.nombre)
                eFechaPlanificacionSedeVirtualExamenes = FechaPlanificacionSedeVirtualExamen.objects.filter(
                    sede_id=sede_id,
                    periodo_id=periodo_id,
                    fecha__in=fechas).order_by(
                    'fecha')
                for eFechaPlanificacionSedeVirtualExamen in eFechaPlanificacionSedeVirtualExamenes:
                    fecha = eFechaPlanificacionSedeVirtualExamen.fecha
                    eTurnoPlanificacionSedeVirtualExamenes = TurnoPlanificacionSedeVirtualExamen.objects.filter(
                        fechaplanificacion=eFechaPlanificacionSedeVirtualExamen).order_by('horainicio')
                    for eTurnoPlanificacionSedeVirtualExamen in eTurnoPlanificacionSedeVirtualExamenes:
                        horainicio = eTurnoPlanificacionSedeVirtualExamen.horainicio
                        horafin = eTurnoPlanificacionSedeVirtualExamen.horafin

                        # eAulaPlanificacionSedeVirtualExamenes = AulaPlanificacionSedeVirtualExamen.objects.filter(turnoplanificacion__fechaplanificacion__sede_id=sede_id,
                        #                                                                                           turnoplanificacion__fechaplanificacion__periodo_id=periodo_id)
                        eAulaPlanificacionSedeVirtualExamenes = AulaPlanificacionSedeVirtualExamen.objects.filter(
                            turnoplanificacion=eTurnoPlanificacionSedeVirtualExamen)
                        eAulaPlanificacionSedeVirtualExamenes_exclude_llenos = eAulaPlanificacionSedeVirtualExamenes.annotate(
                            total_general=Count('materiaasignadaplanificacionsedevirtualexamen__id', filter=Q(
                                materiaasignadaplanificacionsedevirtualexamen__materiaasignada__matricula__nivel__periodo_id=periodo_id,
                                status=True))).filter(total_general=F('aula__capacidad'))
                        eAulaPlanificacionSedeVirtualExamenes.exclude(
                            pk__in=eAulaPlanificacionSedeVirtualExamenes_exclude_llenos.values_list("id", flat=True))
                        eAulaPlanificacionSedeVirtualExamenes = eAulaPlanificacionSedeVirtualExamenes.order_by(
                            'turnoplanificacion__fechaplanificacion__fecha',
                            'turnoplanificacion__horainicio').distinct()
                        # totalAulaSinLlenar = len(eAulaPlanificacionSedeVirtualExamenes.values("id"))
                        contadorAulaSinLlenar = 0
                        # banderaBreakAula = False
                        for eAulaPlanificacionSedeVirtualExamen in eAulaPlanificacionSedeVirtualExamenes:
                            eAula = eAulaPlanificacionSedeVirtualExamen.aula
                            capacidad = eAula.capacidad
                            cantidadad_planificadas = eAulaPlanificacionSedeVirtualExamen.cantidadad_planificadas()
                            # eTurnoPlanificacionSedeVirtualExamen = eAulaPlanificacionSedeVirtualExamen.turnoplanificacion
                            # horainicio = eTurnoPlanificacionSedeVirtualExamen.horainicio
                            # horafin = eTurnoPlanificacionSedeVirtualExamen.horafin
                            # eFechaPlanificacionSedeVirtualExamen = eTurnoPlanificacionSedeVirtualExamen.fechaplanificacion
                            # fecha = eFechaPlanificacionSedeVirtualExamen.fecha
                            if cantidadad_planificadas < capacidad:
                                print(
                                    f"Se procede a planificar en la fecha {fecha} en el horario {horainicio} a {horafin} en el aula {eAula.nombre}")
                                eMateriaAsignadaPlanificacionSedeVirtualExamenes = MateriaAsignadaPlanificacionSedeVirtualExamen.objects.filter(
                                    status=True,
                                    aulaplanificacion__turnoplanificacion__fechaplanificacion__periodo_id=periodo_id,
                                    detallemodeloevaluativo_id=detallemodeloevaluativo_id).exclude(
                                    materiaasignada__materia_id__in=materias_no_planifica).values_list(
                                    "materiaasignada__id",
                                    flat=True)
                                eMateriaAsignadaPlanificacionSedeVirtualExamenesmatriculas = MateriaAsignadaPlanificacionSedeVirtualExamen.objects.filter(
                                    status=True,
                                    aulaplanificacion__turnoplanificacion__fechaplanificacion__periodo_id=periodo_id,
                                    detallemodeloevaluativo_id=detallemodeloevaluativo_id).exclude(
                                    materiaasignada__materia_id__in=materias_no_planifica).values_list(
                                    "materiaasignada__matricula__id", flat=True)

                                filter_conflicto = (Q(aulaplanificacion__turnoplanificacion__horainicio__lte=horafin,
                                                      aulaplanificacion__turnoplanificacion__horafin__gte=horafin,
                                                      aulaplanificacion__turnoplanificacion__fechaplanificacion__fecha=fecha) |
                                                    Q(aulaplanificacion__turnoplanificacion__horainicio__lte=horainicio,
                                                      aulaplanificacion__turnoplanificacion__horafin__gte=horainicio,
                                                      aulaplanificacion__turnoplanificacion__fechaplanificacion__fecha=fecha))
                                eMatriculas = Matricula.objects.filter(
                                    pk__in=eMatriculaSedeExamenes.values_list("matricula__id", flat=True),
                                    status=True,
                                    retiradomatricula=False,
                                    bloqueomatricula=False,
                                    nivel__periodo_id=periodo_id,
                                    inscripcion__inscripcionmalla__malla=eMalla).exclude(
                                    id__in=materias_completas).filter(
                                    materiaasignada__materia__asignaturamalla__nivelmalla_id=9)
                                # eMatriculas = eMatriculas.filter(pk__in=ids_matricula)
                                eMatriculas_exclude_planificadas = eMatriculas.annotate(
                                    total_planificadas=Count('materiaasignada__id', filter=Q(
                                        materiaasignada__id__in=eMateriaAsignadaPlanificacionSedeVirtualExamenes,
                                        status=True),
                                                             nivel__periodo_id=periodo_id, status=True),
                                    total_general=Count('materiaasignada__id',
                                                        filter=Q(nivel__periodo_id=periodo_id, status=True),
                                                        exclude=Q(materiaasignada__materia__asignatura__id=4837) | Q(
                                                            materiaasignada__materia__asignaturamalla__malla_id__in=eMallasIngles.values_list(
                                                                "id", flat=True)))).filter(
                                    Q(total_general=F('total_planificadas')))
                                eMatriculas_exclude_planificadas_x_dia = eMatriculas.annotate(
                                    total_planificadas=Count('materiaasignada__id', filter=Q(
                                        materiaasignada__id__in=eMateriaAsignadaPlanificacionSedeVirtualExamenes.filter(
                                            aulaplanificacion__turnoplanificacion__fechaplanificacion__fecha=fecha),
                                        status=True), nivel__periodo_id=periodo_id, status=True)).filter(
                                    Q(total_planificadas=limite_x_día))
                                eMatriculas_exclude_ingles = eMatriculas.annotate(
                                    total_ingles=Count('materiaasignada__materia__asignaturamalla__id', filter=Q(
                                        materiaasignada__materia__asignaturamalla__malla_id__in=eMallasIngles.values_list(
                                            "id",
                                            flat=True),
                                        nivel__periodo_id=periodo_id, status=True)),
                                    total_general=Count('materiaasignada__materia__asignaturamalla__id',
                                                        filter=Q(nivel__periodo_id=periodo_id, status=True))).filter(
                                    total_general=F('total_ingles'))
                                ids_exclude = list(
                                    eMateriaAsignadaPlanificacionSedeVirtualExamenesmatriculas.filter(filter_conflicto))
                                ids_exclude.extend(list(eMatriculas_exclude_planificadas.values_list('id', flat=True)))
                                ids_exclude.extend(
                                    list(eMatriculas_exclude_planificadas_x_dia.values_list('id', flat=True)))
                                ids_exclude.extend(list(eMatriculas_exclude_ingles.values_list('id', flat=True)))
                                sql = f"""SET statement_timeout='20000 s';
                                SELECT
                                            "sga_matricula"."id",
                                            COUNT("sga_materia"."asignaturamalla_id")
                                                    FILTER (WHERE ("sga_asignaturamalla"."malla_id" NOT IN (	SELECT U0."id"
                                                                                                                            FROM "sga_malla" U0
                                                                                                                            WHERE U0."id" IN (353, 22)
                                                                                                                        ) AND
                                                                        "sga_nivel"."periodo_id" = {periodo_id} AND
                                                                        "sga_matricula"."status"
                                                                        )
                                                            ) AS "total_general",
                                            COUNT("sga_materia"."asignaturamalla_id")
                                                    FILTER (WHERE ("sga_materiaasignada"."id" IN (	SELECT U0."materiaasignada_id"
                                                                                                                            FROM "inno_materiaasignadaplanificacionsedevirtualexamen" U0
                                                                                                                            INNER JOIN "sga_materiaasignada" U1 ON U1."id" = U0.materiaasignada_id
                                                                                                                            WHERE U1."matricula_id" = "sga_matricula"."id"
                                                                                                                        ) AND
                                                                        "sga_nivel"."periodo_id" = {periodo_id} AND
                                                                        "sga_matricula"."status")
                                                            ) AS "total_planificadas"
                                        FROM "sga_matricula"
                                        INNER JOIN "sga_inscripcion" ON "sga_matricula"."inscripcion_id" = "sga_inscripcion"."id"
                                        INNER JOIN "sga_inscripcionmalla" ON "sga_inscripcion"."id" = "sga_inscripcionmalla"."inscripcion_id"
                                        INNER JOIN "sga_nivel" ON "sga_matricula"."nivel_id" = "sga_nivel"."id"
                                        INNER JOIN "sga_periodo" ON "sga_nivel"."periodo_id" = "sga_periodo"."id"
                                        LEFT OUTER JOIN "sga_materiaasignada" ON "sga_matricula"."id" = "sga_materiaasignada"."matricula_id"
                                        LEFT OUTER JOIN "sga_materia" ON "sga_materiaasignada"."materia_id" = "sga_materia"."id"
                                        LEFT OUTER JOIN "sga_asignaturamalla" ON "sga_materia"."asignaturamalla_id" = "sga_asignaturamalla"."id"
                                        WHERE (
                                            NOT "sga_matricula"."bloqueomatricula" AND
                                            "sga_inscripcionmalla"."malla_id" = {eMalla.pk} AND
                                            "sga_nivel"."periodo_id" = {periodo_id} AND
                                            "sga_matricula"."id" IN (
                                                                                SELECT DISTINCT
                                                                                    U0."matricula_id"
                                                                                FROM "inno_matriculasedeexamen" U0
                                                                                    INNER JOIN "sga_matricula" U2 ON U0."matricula_id" = U2."id"
                                                                                    INNER JOIN "sga_nivel" U3 ON U2."nivel_id" = U3."id"
                                                                                WHERE (
                                                                                            U0."detallemodeloevaluativo_id" = {detallemodeloevaluativo_id} AND
                                                                                            U3."periodo_id" = {periodo_id} AND
                                                                                            NOT U2."retiradomatricula" AND
                                                                                            U2."status" AND
                                                                                            U0."sede_id" = {sede_id} AND
                                                                                            U0."status"
                                                                                        )
                                                                            ) AND
                                            NOT "sga_matricula"."retiradomatricula" AND
                                            "sga_matricula"."status"
                                            )
                                        GROUP BY "sga_matricula"."id"
                                        HAVING
                                                COUNT("sga_materia"."asignaturamalla_id")
                                                    FILTER (WHERE ("sga_asignaturamalla"."malla_id" NOT IN (	SELECT U0."id"
                                                                                                                            FROM "sga_malla" U0
                                                                                                                            WHERE U0."id" IN (353, 22)
                                                                                                                        ) AND
                                                                        "sga_nivel"."periodo_id" = {periodo_id} AND
                                                                        "sga_matricula"."status"
                                                                        )
                                                            )
                                                <>
                                                COUNT("sga_materia"."asignaturamalla_id")
                                                    FILTER (WHERE ("sga_materiaasignada"."id" IN (	SELECT U0."materiaasignada_id"
                                                                                                                            FROM "inno_materiaasignadaplanificacionsedevirtualexamen" U0
                                                                                                                            INNER JOIN "sga_materiaasignada" U1 ON U1."id" = U0.materiaasignada_id
                                                                                                                            WHERE U1."matricula_id" = "sga_matricula"."id"
                                                                                                                        ) AND
                                                                        "sga_nivel"."periodo_id" = {periodo_id} AND
                                                                        "sga_matricula"."status")
                                        )"""
                                eMatriculas = eMatriculas.exclude(pk__in=ids_exclude)
                                cursor.execute(sql)
                                results = cursor.fetchall()
                                ids_matricula = [r[0] for r in results]
                                eMatriculas = eMatriculas.filter(pk__in=ids_matricula)
                                eMatriculas = eMatriculas.order_by('inscripcion__inscripcionnivel__nivel__orden',
                                                                   'inscripcion__persona__apellido1',
                                                                   'inscripcion__persona__apellido2',
                                                                   'inscripcion__persona__nombres').distinct()
                                contador = cantidadad_planificadas
                                if not eMatriculas.values("id").exists():
                                    contadorAulaSinLlenar += 1

                                if contadorAulaSinLlenar > 0:
                                    # banderaBreakAula = True
                                    break
                                for eMatricula in eMatriculas:
                                    matricula_id = eMatricula.id
                                    # mat = MatriculaSedeExamen.objects.filter(status=True, matricula=eMatricula,
                                    #                                          sede_id=sede_id,
                                    #                                          detallemodeloevaluativo_id=detallemodeloevaluativo_id).first()
                                    delet = MateriaAsignadaPlanificacionSedeVirtualExamen.objects.filter(status=True,
                                                                                                         materiaasignada__matricula_id=matricula_id).exclude(
                                        Q(materiaasignada__materia__asignatura__id=4837) |
                                        Q(materiaasignada__materia__asignaturamalla__malla_id__in=eMallasIngles.values_list(
                                            "id", flat=True))).exclude(
                                        materiaasignada__materia_id__in=materias_no_planifica).values('id').filter(
                                        materiaasignada__materia__asignaturamalla__nivelmalla_id=9)
                                    tmaterias = MateriaAsignada.objects.filter(status=True,
                                                                               matricula_id=matricula_id).exclude(
                                        Q(materia__asignatura__id=4837) |
                                        Q(materia__asignaturamalla__malla_id__in=eMallasIngles.values_list("id",
                                                                                                           flat=True))).exclude(
                                        materia_id__in=materias_no_planifica).values(
                                        'id').exclude(materia__asignaturamalla__nivelmalla_id=9)
                                    # if len(tmaterias) < 5 or len(tmaterias) > 6:
                                    #     continue
                                    if len(delet) == len(tmaterias) and not matricula_id in materias_completas:
                                        materias_completas.append(matricula_id)
                                    eMateriaAsignadas = MateriaAsignada.objects.filter(status=True,
                                                                                       matricula=eMatricula).exclude(
                                        materia_id__in=materias_no_planifica).exclude(
                                        materia__asignaturamalla__nivelmalla_id=9)
                                    eMateriaAsignadas = eMateriaAsignadas.exclude(
                                        materia__asignaturamalla__nivelmalla_id=9).exclude(
                                        Q(pk__in=eMateriaAsignadaPlanificacionSedeVirtualExamenes) |
                                        Q(materia__asignatura__id=4837) |
                                        Q(materia__asignaturamalla__malla_id__in=eMallasIngles.values_list("id",
                                                                                                           flat=True)))
                                    eMateriaAsignadas = eMateriaAsignadas.order_by(
                                        'materia__asignaturamalla__nivelmalla__orden')
                                    if eMateriaAsignadas.values("id").exists():
                                        eMateriaAsignada = eMateriaAsignadas.first()
                                        if eMateriaAsignada.materia.asignaturamalla.transversal or eMateriaAsignada.materia.modeloevaluativo_id == 27:
                                            detallemodeloevaluativo_id = 123
                                        contador += 1
                                        eMateriaAsignadaPlanificacionSedeVirtualExamen = MateriaAsignadaPlanificacionSedeVirtualExamen(
                                            aulaplanificacion=eAulaPlanificacionSedeVirtualExamen,
                                            materiaasignada=eMateriaAsignada,
                                            detallemodeloevaluativo_id=detallemodeloevaluativo_id,
                                            utilizar_qr=True
                                        )
                                        print(
                                            f"------- ({contador}) Se asignada el estudiante {eMateriaAsignada.matricula.inscripcion.persona} en la asignatura {eMateriaAsignada.materia.asignatura.nombre}")
                                        eMateriaAsignadaPlanificacionSedeVirtualExamen.save()
                                        # if mat:
                                        #     if not mat.aceptotermino:
                                        #         result = generate_qr_examen_final(
                                        #             eMateriaAsignadaPlanificacionSedeVirtualExamen,
                                        #             materiaasignada_id=eMateriaAsignadaPlanificacionSedeVirtualExamen.materiaasignada_id)
                                        #
                                        # eMateriaAsignada.save()
                                        # for mate in MateriaAsignada.objects.filter(status=True, matricula=eMatricula):
                                        #     mate.visiblehorarioexamen = False
                                        #     mate.save()
                                        if contador >= capacidad:
                                            break
        except Exception as ex:
            print(ex)


def horarios_9_niveles():
    for materia in Materia.objects.filter(nivel__periodo_id=224, asignaturamalla__nivelmalla__gte=9, asignaturamalla__malla__carrera__modalidad=3):
        print(materia)
        for horario in materia.horarioexamen_set.filter(status=True, detallemodelo_id__in=[123]):
            fechaplan = FechaPlanificacionSedeVirtualExamen.objects.filter(sede_id=1, periodo_id=224, fecha=horario.fecha).first()
            if not fechaplan:
                fechaplan= FechaPlanificacionSedeVirtualExamen(sede_id=1, periodo_id=224, fecha=horario.fecha)
                fechaplan.save()
            print(fechaplan.fecha)
            detalle = horario.detallehoraexamen().first()
            if detalle:
                turnoplan = TurnoPlanificacionSedeVirtualExamen.objects.filter(fechaplanificacion=fechaplan, horainicio=detalle.horainicio, horafin=detalle.horafin).first()
                if not turnoplan:
                    turnoplan = TurnoPlanificacionSedeVirtualExamen(fechaplanificacion=fechaplan, horainicio=detalle.horainicio, horafin=detalle.horafin, horas=1)
                    turnoplan.save()
                print(turnoplan.horainicio, ' - ', turnoplan.horafin)
                aulaplan = AulaPlanificacionSedeVirtualExamen.objects.filter(turnoplanificacion=turnoplan, aula_id=285).first()
                if not aulaplan:
                    aulaplan = AulaPlanificacionSedeVirtualExamen(turnoplanificacion=turnoplan, aula_id=285)
                    aulaplan.save()
                print(aulaplan.aula)
                for asignado in materia.materiaasignada_set.filter(status=True, retiramateria=False):
                    eMateriaAsignadaPlanificacionSedeVirtualExamen = MateriaAsignadaPlanificacionSedeVirtualExamen.objects.filter(status=True, materiaasignada=asignado, detallemodeloevaluativo_id=123).first()
                    print(asignado.matricula.inscripcion)
                    if not eMateriaAsignadaPlanificacionSedeVirtualExamen:
                        print('PLANIFICANDO MATERIA..............')
                        eMateriaAsignadaPlanificacionSedeVirtualExamen = MateriaAsignadaPlanificacionSedeVirtualExamen(
                        aulaplanificacion=aulaplan, materiaasignada=asignado, detallemodeloevaluativo_id=123, utilizar_qr=True)
                        eMateriaAsignadaPlanificacionSedeVirtualExamen.save()
                    else:
                        if eMateriaAsignadaPlanificacionSedeVirtualExamen.aulaplanificacion==aulaplan:
                            print('-----> MATERIA PLANIFICADA')
                        else:
                            eMateriaAsignadaPlanificacionSedeVirtualExamen.aulaplanificacion=aulaplan
                            eMateriaAsignadaPlanificacionSedeVirtualExamen.save()
                            print('*****CAMBIANDO PLANIFICACION DE MATERIA****')
            else:
                print('SIN EXAMENEN PLANIFICADO')
    # planificar_pregrado_unemi_2023(4, 1, 224, 37, fechas_examenes_2023_milagro)

# horarios_9_niveles()


def dividir_turnos2(limite):
    try:
        for fecha in FechaPlanificacionSedeVirtualExamen.objects.filter(status=True, sede_id=11, fecha__gte='2024-01-16', periodo_id=224):
            for turno in fecha.turnoplanificacionsedevirtualexamen_set.filter(status=True).exclude(id__in=[2934, 2789]):
                for eLaboratorioVirtual in LaboratorioVirtual.objects.filter(sedevirtual_id=11)[:120]:
                    eLaboratorioVirtual.capacidad = limite
                    eLaboratorioVirtual.save(update_fields=['capacidad'])
                    eAulaPlanificacionSedeVirtualExamenes = AulaPlanificacionSedeVirtualExamen.objects.filter(
                        turnoplanificacion=turno, aula=eLaboratorioVirtual)
                    if not eAulaPlanificacionSedeVirtualExamenes.values("id").exists():
                        eAulaPlanificacionSedeVirtualExamen = AulaPlanificacionSedeVirtualExamen(
                            turnoplanificacion=turno,
                            aula=eLaboratorioVirtual)
                        eAulaPlanificacionSedeVirtualExamen.save()
                for aula in turno.aulaplanificacionsedevirtualexamen_set.filter(status=True).exclude(aula_id__in=(285, 287)):
                    c = 0
                    restante = aula.cantidadad_planificadas() - limite
                    if restante == 0:
                        continue
                    for alu in aula.materiaasignadaplanificadas():
                        if aula.cantidadad_planificadas() - limite <= 0:
                            break
                        c += 1
                        alu.save()
                        if c >= limite:
                            for aula_m in turno.aulaplanificacionsedevirtualexamen_set.filter(status=True).exclude(aula_id__in=(285, 287)):
                                if aula.cantidadad_planificadas() - limite <= 0:
                                    break
                                if aula_m.cantidadad_planificadas() >= limite:
                                    continue
                                if restante < limite:
                                    aux = restante
                                    for materiaplan in aula.materiaasignadaplanificadas()[:restante]:
                                        materiaplan.aulaplanificacion = aula_m
                                        materiaplan.save()
                                        print(materiaplan)
                                        aux-=1
                                        if aula_m.cantidadad_planificadas() >= limite:
                                            restante = aux
                                            break


    except Exception as ex:
        print(ex)


def dividir_turnos(limite):
    try:
        for fecha in FechaPlanificacionSedeVirtualExamen.objects.filter(status=True, sede_id=11, fecha__gte='2024-01-16', periodo_id=224):
            for turno in fecha.turnoplanificacionsedevirtualexamen_set.filter(status=True):
                for eLaboratorioVirtual in LaboratorioVirtual.objects.filter(sedevirtual_id=11)[:95]:
                    eLaboratorioVirtual.capacidad = limite
                    eLaboratorioVirtual.save(update_fields=['capacidad'])
                    eAulaPlanificacionSedeVirtualExamenes = AulaPlanificacionSedeVirtualExamen.objects.filter(
                        turnoplanificacion=turno, aula=eLaboratorioVirtual)
                    if not eAulaPlanificacionSedeVirtualExamenes.values("id").exists():
                        eAulaPlanificacionSedeVirtualExamen = AulaPlanificacionSedeVirtualExamen(
                            turnoplanificacion=turno,
                            aula=eLaboratorioVirtual)
                        eAulaPlanificacionSedeVirtualExamen.save()
                for aula in turno.aulaplanificacionsedevirtualexamen_set.filter(status=True, aula_id__in=(285, 287)):
                    for aula_n in turno.aulaplanificacionsedevirtualexamen_set.filter(status=True).exclude(aula_id__in=(285, 287)):
                        if aula_n.cantidadad_planificadas() >= limite:
                            continue
                        for materiaplan in aula.materiaasignadaplanificadas():
                            materiaplan.aulaplanificacion=aula_n
                            materiaplan.save(update_fields=['aulaplanificacion'])
                            print(materiaplan)
                            if aula_n.cantidadad_planificadas() >= limite:
                                break
        dividir_turnos2(40)
    except Exception as ex:
        print(ex)


import concurrent.futures


def funcion_global(n):
    print("Función {} ejecutada".format(n))
    print("Este es un numero al azar de la funcion {} .... {}".format(n, random.randint(1, 10000000)))
    return True



def funcion_2():
    # Código de la función 2
    print("Función 2 ejecutada")
    return 2


def funcion_3():
    # Código de la función 3
    print("Función 3 ejecutada")
    return 3

fechas_examenes_2023_milagro = [
    (datetime(2024, 1, 23, 0, 0, 0)).date(),
    (datetime(2024, 1, 24, 0, 0, 0)).date()
]

def planificar_pregrado_unemi_202(limite_x_día, sede, periodo_id, detallemodeloevaluativo_id, fechas, malla_id):
    try:
        sede_id = sede
        print('PLANIFICANDO SEDE ' + str(sede_id))
        # periodo_id = 153
        # detallemodeloevaluativo_id = 37
        materias_completas = []
        cursor = connections['sga_select'].cursor()
        eMallasIngles = Malla.objects.filter(pk__in=[353, 22]).values_list('id', flat=True)
        eMallasMedicina = Malla.objects.filter(carrera__id__in=[223]).values_list('id', flat=True)
        eMatriculaSedeExamenes = MatriculaSedeExamen.objects.filter(detallemodeloevaluativo_id=detallemodeloevaluativo_id,
                                                                    status=True, matricula__status=True,
                                                                    matricula__retiradomatricula=False,
                                                                    matricula__nivel__periodo_id=periodo_id,
                                                                    matricula__inscripcion__carrera__modalidad=3).filter(matricula__materiaasignada__materia__asignaturamalla__nivelmalla_id=9)
        # if DEBUG:
        # MateriaAsignadaPlanificacionSedeVirtualExamen.objects.filter(status=True,
        #                                                              aulaplanificacion__turnoplanificacion__fechaplanificacion__periodo_id=periodo_id,
        #                                                              aulaplanificacion__turnoplanificacion__fechaplanificacion__sede_id=sede_id,
        #                                                              detallemodeloevaluativo_id=detallemodeloevaluativo_id).delete()
        # eMatriculaSedeExamenes.update(sede_id=sede_id)

        eMatriculas = Matricula.objects.filter(pk__in=eMatriculaSedeExamenes.values_list("matricula__id", flat=True),
                                               status=True, retiradomatricula=False,
                                               nivel__periodo_id=periodo_id).filter(materiaasignada__materia__asignaturamalla__nivelmalla_id=9)
        eMatriculas_exclude_ingles = eMatriculas.annotate(
            total_ingles=Count('materiaasignada__materia__asignaturamalla__id', filter=Q(
                materiaasignada__materia__asignaturamalla__malla_id__in=eMallasIngles.values_list('id', flat=True),
                nivel__periodo_id=periodo_id, status=True)),
            total_general=Count('materiaasignada__materia__asignaturamalla__id',
                                filter=Q(nivel__periodo_id=periodo_id, status=True))).filter(
            total_general=F('total_ingles'))
        ids_exclude = []
        ids_exclude.extend(list(eMatriculas_exclude_ingles.values_list('id', flat=True)))
        eMatriculas = eMatriculas.exclude(pk__in=ids_exclude)
        eMallas = Malla.objects.filter(carrera__modalidad=3, id=malla_id,
                                       pk__in=eMatriculas.values_list('inscripcion__inscripcionmalla__malla_id',
                                                                      flat=True).distinct()).exclude(
            id__in=eMallasIngles)
        # eCarreras = Carrera.objects.filter(pk__in=eMallas.values_list('carrera__id', flat=True))
        # eCoordinaciones = Coordinacion.objects.filter(pk__in=eCarreras.values_list('coordinacion__id', flat=True)).distinct()
        # for eCoordinacion in eCoordinaciones:
        #     eMallas = eMallas.filter(carrera_id__in=eCoordinacion.carreras().values_list('id', flat=True)).order_by('carrera__nombre', 'inicio')
        eMallas = eMallas.order_by('carrera__nombre', 'inicio')
        for eMalla in eMallas:
            print(eMalla.carrera.nombre)
            eFechaPlanificacionSedeVirtualExamenes = FechaPlanificacionSedeVirtualExamen.objects.filter(sede_id=sede_id,
                                                                                                        periodo_id=periodo_id,
                                                                                                        fecha__in=fechas).order_by(
                'fecha')
            for eFechaPlanificacionSedeVirtualExamen in eFechaPlanificacionSedeVirtualExamenes:
                fecha = eFechaPlanificacionSedeVirtualExamen.fecha
                eTurnoPlanificacionSedeVirtualExamenes = TurnoPlanificacionSedeVirtualExamen.objects.filter(
                    fechaplanificacion=eFechaPlanificacionSedeVirtualExamen).order_by('horainicio')
                for eTurnoPlanificacionSedeVirtualExamen in eTurnoPlanificacionSedeVirtualExamenes:
                    horainicio = eTurnoPlanificacionSedeVirtualExamen.horainicio
                    horafin = eTurnoPlanificacionSedeVirtualExamen.horafin

                    # eAulaPlanificacionSedeVirtualExamenes = AulaPlanificacionSedeVirtualExamen.objects.filter(turnoplanificacion__fechaplanificacion__sede_id=sede_id,
                    #                                                                                           turnoplanificacion__fechaplanificacion__periodo_id=periodo_id)
                    eAulaPlanificacionSedeVirtualExamenes = AulaPlanificacionSedeVirtualExamen.objects.filter(
                        turnoplanificacion=eTurnoPlanificacionSedeVirtualExamen)
                    eAulaPlanificacionSedeVirtualExamenes_exclude_llenos = eAulaPlanificacionSedeVirtualExamenes.annotate(
                        total_general=Count('materiaasignadaplanificacionsedevirtualexamen__id', filter=Q(
                            materiaasignadaplanificacionsedevirtualexamen__materiaasignada__matricula__nivel__periodo_id=periodo_id,
                            status=True))).filter(total_general=F('aula__capacidad'))
                    eAulaPlanificacionSedeVirtualExamenes.exclude(
                        pk__in=eAulaPlanificacionSedeVirtualExamenes_exclude_llenos.values_list("id", flat=True))
                    eAulaPlanificacionSedeVirtualExamenes = eAulaPlanificacionSedeVirtualExamenes.order_by(
                        'turnoplanificacion__fechaplanificacion__fecha',
                        'turnoplanificacion__horainicio').distinct()
                    # totalAulaSinLlenar = len(eAulaPlanificacionSedeVirtualExamenes.values("id"))
                    contadorAulaSinLlenar = 0
                    # banderaBreakAula = False
                    for eAulaPlanificacionSedeVirtualExamen in eAulaPlanificacionSedeVirtualExamenes:
                        eAula = eAulaPlanificacionSedeVirtualExamen.aula
                        capacidad = eAula.capacidad
                        cantidadad_planificadas = eAulaPlanificacionSedeVirtualExamen.cantidadad_planificadas()
                        # eTurnoPlanificacionSedeVirtualExamen = eAulaPlanificacionSedeVirtualExamen.turnoplanificacion
                        # horainicio = eTurnoPlanificacionSedeVirtualExamen.horainicio
                        # horafin = eTurnoPlanificacionSedeVirtualExamen.horafin
                        # eFechaPlanificacionSedeVirtualExamen = eTurnoPlanificacionSedeVirtualExamen.fechaplanificacion
                        # fecha = eFechaPlanificacionSedeVirtualExamen.fecha
                        if cantidadad_planificadas < capacidad:
                            print(
                                f"Se procede a planificar en la fecha {fecha} en el horario {horainicio} a {horafin} en el aula {eAula.nombre}")
                            eMateriaAsignadaPlanificacionSedeVirtualExamenes = MateriaAsignadaPlanificacionSedeVirtualExamen.objects.filter(
                                status=True,
                                aulaplanificacion__turnoplanificacion__fechaplanificacion__periodo_id=periodo_id,
                                detallemodeloevaluativo_id=detallemodeloevaluativo_id).exclude(
                                materiaasignada__materia_id__in=materias_no_planifica).values_list(
                                "materiaasignada__id",
                                flat=True)
                            eMateriaAsignadaPlanificacionSedeVirtualExamenesmatriculas = MateriaAsignadaPlanificacionSedeVirtualExamen.objects.filter(
                                status=True,
                                aulaplanificacion__turnoplanificacion__fechaplanificacion__periodo_id=periodo_id,
                                detallemodeloevaluativo_id=detallemodeloevaluativo_id).exclude(
                                materiaasignada__materia_id__in=materias_no_planifica).values_list(
                                "materiaasignada__matricula__id", flat=True)

                            filter_conflicto = (Q(aulaplanificacion__turnoplanificacion__horainicio__lte=horafin,
                                                  aulaplanificacion__turnoplanificacion__horafin__gte=horafin,
                                                  aulaplanificacion__turnoplanificacion__fechaplanificacion__fecha=fecha) |
                                                Q(aulaplanificacion__turnoplanificacion__horainicio__lte=horainicio,
                                                  aulaplanificacion__turnoplanificacion__horafin__gte=horainicio,
                                                  aulaplanificacion__turnoplanificacion__fechaplanificacion__fecha=fecha))
                            eMatriculas = Matricula.objects.filter(
                                pk__in=eMatriculaSedeExamenes.values_list("matricula__id", flat=True),
                                status=True,
                                retiradomatricula=False,
                                bloqueomatricula=False,
                                nivel__periodo_id=periodo_id,
                                inscripcion__inscripcionmalla__malla=eMalla).exclude(id__in=materias_completas).filter(materiaasignada__materia__asignaturamalla__nivelmalla_id=9)
                            # eMatriculas = eMatriculas.filter(pk__in=ids_matricula)
                            eMatriculas_exclude_planificadas = eMatriculas.annotate(
                                total_planificadas=Count('materiaasignada__id', filter=Q(
                                    materiaasignada__id__in=eMateriaAsignadaPlanificacionSedeVirtualExamenes,
                                    status=True),
                                                         nivel__periodo_id=periodo_id, status=True),
                                total_general=Count('materiaasignada__id',
                                                    filter=Q(nivel__periodo_id=periodo_id, status=True),
                                                    exclude=Q(materiaasignada__materia__asignatura__id=4837) | Q(
                                                        materiaasignada__materia__asignaturamalla__malla_id__in=eMallasIngles.values_list(
                                                            "id", flat=True)))).filter(
                                Q(total_general=F('total_planificadas')))
                            eMatriculas_exclude_planificadas_x_dia = eMatriculas.annotate(
                                total_planificadas=Count('materiaasignada__id', filter=Q(
                                    materiaasignada__id__in=eMateriaAsignadaPlanificacionSedeVirtualExamenes.filter(
                                        aulaplanificacion__turnoplanificacion__fechaplanificacion__fecha=fecha),
                                    status=True), nivel__periodo_id=periodo_id, status=True)).filter(
                                Q(total_planificadas=limite_x_día))
                            eMatriculas_exclude_ingles = eMatriculas.annotate(
                                total_ingles=Count('materiaasignada__materia__asignaturamalla__id', filter=Q(
                                    materiaasignada__materia__asignaturamalla__malla_id__in=eMallasIngles.values_list(
                                        "id",
                                        flat=True),
                                    nivel__periodo_id=periodo_id, status=True)),
                                total_general=Count('materiaasignada__materia__asignaturamalla__id',
                                                    filter=Q(nivel__periodo_id=periodo_id, status=True))).filter(
                                total_general=F('total_ingles'))
                            ids_exclude = list(
                                eMateriaAsignadaPlanificacionSedeVirtualExamenesmatriculas.filter(filter_conflicto))
                            ids_exclude.extend(list(eMatriculas_exclude_planificadas.values_list('id', flat=True)))
                            ids_exclude.extend(
                                list(eMatriculas_exclude_planificadas_x_dia.values_list('id', flat=True)))
                            ids_exclude.extend(list(eMatriculas_exclude_ingles.values_list('id', flat=True)))
                            sql = f"""SET statement_timeout='20000 s';
                            SELECT
                                        "sga_matricula"."id",
                                        COUNT("sga_materia"."asignaturamalla_id")
                                                FILTER (WHERE ("sga_asignaturamalla"."malla_id" NOT IN (	SELECT U0."id"
                                                                                                                        FROM "sga_malla" U0
                                                                                                                        WHERE U0."id" IN (353, 22)
                                                                                                                    ) AND
                                                                    "sga_nivel"."periodo_id" = {periodo_id} AND
                                                                    "sga_matricula"."status"
                                                                    )
                                                        ) AS "total_general",
                                        COUNT("sga_materia"."asignaturamalla_id")
                                                FILTER (WHERE ("sga_materiaasignada"."id" IN (	SELECT U0."materiaasignada_id"
                                                                                                                        FROM "inno_materiaasignadaplanificacionsedevirtualexamen" U0
                                                                                                                        INNER JOIN "sga_materiaasignada" U1 ON U1."id" = U0.materiaasignada_id
                                                                                                                        WHERE U1."matricula_id" = "sga_matricula"."id"
                                                                                                                    ) AND
                                                                    "sga_nivel"."periodo_id" = {periodo_id} AND
                                                                    "sga_matricula"."status")
                                                        ) AS "total_planificadas"
                                    FROM "sga_matricula"
                                    INNER JOIN "sga_inscripcion" ON "sga_matricula"."inscripcion_id" = "sga_inscripcion"."id"
                                    INNER JOIN "sga_inscripcionmalla" ON "sga_inscripcion"."id" = "sga_inscripcionmalla"."inscripcion_id"
                                    INNER JOIN "sga_nivel" ON "sga_matricula"."nivel_id" = "sga_nivel"."id"
                                    INNER JOIN "sga_periodo" ON "sga_nivel"."periodo_id" = "sga_periodo"."id"
                                    LEFT OUTER JOIN "sga_materiaasignada" ON "sga_matricula"."id" = "sga_materiaasignada"."matricula_id"
                                    LEFT OUTER JOIN "sga_materia" ON "sga_materiaasignada"."materia_id" = "sga_materia"."id"
                                    LEFT OUTER JOIN "sga_asignaturamalla" ON "sga_materia"."asignaturamalla_id" = "sga_asignaturamalla"."id"
                                    WHERE (
                                        NOT "sga_matricula"."bloqueomatricula" AND
                                        "sga_inscripcionmalla"."malla_id" = {eMalla.pk} AND
                                        "sga_nivel"."periodo_id" = {periodo_id} AND
                                        "sga_matricula"."id" IN (
                                                                            SELECT DISTINCT
                                                                                U0."matricula_id"
                                                                            FROM "inno_matriculasedeexamen" U0
                                                                                INNER JOIN "sga_matricula" U2 ON U0."matricula_id" = U2."id"
                                                                                INNER JOIN "sga_nivel" U3 ON U2."nivel_id" = U3."id"
                                                                            WHERE (
                                                                                        U0."detallemodeloevaluativo_id" = {detallemodeloevaluativo_id} AND
                                                                                        U3."periodo_id" = {periodo_id} AND
                                                                                        NOT U2."retiradomatricula" AND
                                                                                        U2."status" AND
                                                                                        U0."sede_id" = {sede_id} AND
                                                                                        U0."status"
                                                                                    )
                                                                        ) AND
                                        NOT "sga_matricula"."retiradomatricula" AND
                                        "sga_matricula"."status"
                                        )
                                    GROUP BY "sga_matricula"."id"
                                    HAVING
                                            COUNT("sga_materia"."asignaturamalla_id")
                                                FILTER (WHERE ("sga_asignaturamalla"."malla_id" NOT IN (	SELECT U0."id"
                                                                                                                        FROM "sga_malla" U0
                                                                                                                        WHERE U0."id" IN (353, 22)
                                                                                                                    ) AND
                                                                    "sga_nivel"."periodo_id" = {periodo_id} AND
                                                                    "sga_matricula"."status"
                                                                    )
                                                        )
                                            <>
                                            COUNT("sga_materia"."asignaturamalla_id")
                                                FILTER (WHERE ("sga_materiaasignada"."id" IN (	SELECT U0."materiaasignada_id"
                                                                                                                        FROM "inno_materiaasignadaplanificacionsedevirtualexamen" U0
                                                                                                                        INNER JOIN "sga_materiaasignada" U1 ON U1."id" = U0.materiaasignada_id
                                                                                                                        WHERE U1."matricula_id" = "sga_matricula"."id"
                                                                                                                    ) AND
                                                                    "sga_nivel"."periodo_id" = {periodo_id} AND
                                                                    "sga_matricula"."status")
                                    )"""
                            eMatriculas = eMatriculas.exclude(pk__in=ids_exclude).exclude(materiaasignada__materia__asignaturamalla__nivelmalla_id__lte=7)
                            cursor.execute(sql)
                            results = cursor.fetchall()
                            ids_matricula = [r[0] for r in results]
                            eMatriculas = eMatriculas.filter(pk__in=ids_matricula)
                            eMatriculas = eMatriculas.order_by('inscripcion__inscripcionnivel__nivel__orden',
                                                               'inscripcion__persona__apellido1',
                                                               'inscripcion__persona__apellido2',
                                                               'inscripcion__persona__nombres').distinct()
                            contador = cantidadad_planificadas
                            if not eMatriculas.values("id").exists():
                                contadorAulaSinLlenar += 1

                            if contadorAulaSinLlenar > 0:
                                # banderaBreakAula = True
                                break
                            for eMatricula in eMatriculas:
                                matricula_id = eMatricula.id
                                # mat = MatriculaSedeExamen.objects.filter(status=True, matricula=eMatricula,
                                #                                          sede_id=sede_id,
                                #                                          detallemodeloevaluativo_id=detallemodeloevaluativo_id).first()
                                delet = MateriaAsignadaPlanificacionSedeVirtualExamen.objects.filter(status=True,
                                                                                                     materiaasignada__matricula_id=matricula_id).exclude(
                                    Q(materiaasignada__materia__asignatura__id=4837) |
                                    Q(materiaasignada__materia__asignaturamalla__malla_id__in=eMallasIngles.values_list(
                                        "id", flat=True))).exclude(
                                    materiaasignada__materia_id__in=materias_no_planifica).values('id').filter(materiaasignada__materia__asignaturamalla__nivelmalla_id=9)
                                tmaterias = MateriaAsignada.objects.filter(status=True,
                                                                           matricula_id=matricula_id).exclude(
                                    Q(materia__asignatura__id=4837) |
                                    Q(materia__asignaturamalla__malla_id__in=eMallasIngles.values_list("id",
                                                                                                       flat=True))).exclude(
                                    materia_id__in=materias_no_planifica).values(
                                    'id').exclude(materia__asignaturamalla__nivelmalla_id=9)
                                # if len(tmaterias) < 5 or len(tmaterias) > 6:
                                #     continue
                                if len(delet) == len(tmaterias) and not matricula_id in materias_completas:
                                    materias_completas.append(matricula_id)
                                eMateriaAsignadas = MateriaAsignada.objects.filter(status=True,
                                                                                   matricula=eMatricula).exclude(
                                    materia_id__in=materias_no_planifica).exclude(materia__asignaturamalla__nivelmalla_id=9)
                                eMateriaAsignadas = eMateriaAsignadas.exclude(materia__asignaturamalla__nivelmalla_id=9).exclude(
                                    Q(pk__in=eMateriaAsignadaPlanificacionSedeVirtualExamenes) |
                                    Q(materia__asignatura__id=4837) |
                                    Q(materia__asignaturamalla__malla_id__in=eMallasIngles.values_list("id",
                                                                                                       flat=True)))
                                eMateriaAsignadas = eMateriaAsignadas.order_by(
                                    'materia__asignaturamalla__nivelmalla__orden')
                                if eMateriaAsignadas.values("id").exists():
                                    eMateriaAsignada = eMateriaAsignadas.first()
                                    if eMateriaAsignada.materia.asignaturamalla.transversal or eMateriaAsignada.materia.modeloevaluativo_id == 27:
                                        detallemodeloevaluativo_id = 123
                                    contador += 1
                                    eMateriaAsignadaPlanificacionSedeVirtualExamen = MateriaAsignadaPlanificacionSedeVirtualExamen(
                                        aulaplanificacion=eAulaPlanificacionSedeVirtualExamen,
                                        materiaasignada=eMateriaAsignada,
                                        detallemodeloevaluativo_id=detallemodeloevaluativo_id,
                                        utilizar_qr=True
                                    )
                                    print(
                                        f"------- ({contador}) Se asignada el estudiante {eMateriaAsignada.matricula.inscripcion.persona} en la asignatura {eMateriaAsignada.materia.asignatura.nombre}")
                                    eMateriaAsignadaPlanificacionSedeVirtualExamen.save()
                                    # if mat:
                                    #     if not mat.aceptotermino:
                                    #         result = generate_qr_examen_final(
                                    #             eMateriaAsignadaPlanificacionSedeVirtualExamen,
                                    #             materiaasignada_id=eMateriaAsignadaPlanificacionSedeVirtualExamen.materiaasignada_id)
                                    #
                                    # eMateriaAsignada.save()
                                    # for mate in MateriaAsignada.objects.filter(status=True, matricula=eMatricula):
                                    #     mate.visiblehorarioexamen = False
                                    #     mate.save()
                                    if contador >= capacidad:
                                        break
    except Exception as ex:
        print(ex)



# noti = Notificacion(cuerpo=f'total {matriculas}',
#                                 titulo='ids de matriculas recalculadas',
#                                 destinatario_id=33490,
#                                 prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
#                                 tipo=2, en_proceso=False)
# noti.save()


# AlumnosPracticas
#
# GruposProfesorMateria
# periodo_id = 336
# materias = Materia.objects.filter(status=True, nivel__periodo_id=periodo_id, practicas=True, asignaturamalla__malla__carrera__coordinacion=1, asignaturamalla__malla__carrera__nombre__icontains='ENFER')
# profesoresmateria = []
# paralelo = 'B4'
# for nivel in NivelMalla.objects.filter(status=True, id=3):
#     materias_p  = []
#     for materia in materias.filter(asignaturamalla__nivelmalla=nivel, paralelo=paralelo):
#         totalgrupo = 0
#         # total = materia.asignados_a_esta_materia().exclude(matriculas__gte=2)
#         # total2 = materia.asignados_a_esta_materia()
#         tipos = ProfesorMateria.objects.values_list("tipoprofesor_id", flat=True).filter(materia=materia, status=True, activo=True)
#         tiposprofesorpractica = TipoProfesor.objects.filter(pk__in=[2])
#         profesoresmateria = materia.profesores_materia_segun_tipoprofesor_pm(tiposprofesorpractica, multiple=True)
#         totalgrupo = profesoresmateria.count()
#         # if 1 < totalgrupo <= 2:
#         if totalgrupo > 2:
#             materias_p.append(materia.id)
#     ids_matriculas = MateriaAsignada.objects.filter(status=True, retiramateria=False, materia_id__in=materias_p).values_list('matricula_id', flat=True).order_by('matricula__inscripcion__persona__apellido1', 'matricula__inscripcion__persona__apellido2', 'matricula__inscripcion__persona__nombres').distinct()
#     ultimo_grupo = []
#     for materia in materias.filter(asignaturamalla__nivelmalla=nivel, paralelo=paralelo, id__in=materias_p):
#         for mat in ids_matriculas:
#             if not materia.materiaasignada_set.filter(matricula_id=mat).exists():
#                 if not mat in ultimo_grupo:
#                     ultimo_grupo.append(mat)
#                     ma = Matricula.objects.get(id=mat)
#                     # print(f"{ma} -- {materia}")
#
#     g = GruposProfesorMateria.objects.filter(profesormateria__materia_id__in=materias.filter(asignaturamalla__nivelmalla=nivel, paralelo=paralelo, id__in=materias_p))
#     # cupomaximo = g.order_by('cupo').first().cupo
#     # cupomaximo = 21
#     # for pm in profesoresmateria:
#     AlumnosPracticaMateria.objects.filter(status=True, grupoprofesor_id__in=g.values('id')).delete()
#     for mate in materias.filter(asignaturamalla__nivelmalla=nivel, paralelo=paralelo, id__in=materias_p):
#         todosfull = MateriaAsignada.objects.filter(status=True, materia=mate)
#         totalgrupo = GruposProfesorMateria.objects.filter(status=True, profesormateria__materia=mate).order_by('paralelopractica').count()
#         cupomaximo = math.ceil(todosfull.count() / totalgrupo)
#         # cupomaximo = 14
#         c = 0
#         for grupoprofesor in GruposProfesorMateria.objects.filter(status=True, profesormateria__materia=mate).order_by('paralelopractica'):
#             c += 1
#             asignados_otro_grupo = AlumnosPracticaMateria.objects.filter(status=True, materiaasignada__materia=grupoprofesor.profesormateria.materia).values_list('materiaasignada_id', flat=True)
#             alumnos = MateriaAsignada.objects.filter(status=True, materia=grupoprofesor.profesormateria.materia).exclude(id__in=asignados_otro_grupo).order_by('matricula__inscripcion__persona__apellido1', 'matricula__inscripcion__persona__apellido2', 'matricula__inscripcion__persona__nombres')
#             todos = alumnos.values('id')
#
#             # if c >= totalgrupo-1:
#             if c >= totalgrupo:
#                 alumnos = alumnos.filter(Q(matricula_id__in=ultimo_grupo) | Q(id__in=todos))
#             else:
#                 alumnos = alumnos.exclude(matricula_id__in=ultimo_grupo)
#             asignados = AlumnosPracticaMateria.objects.filter(status=True, grupoprofesor=grupoprofesor).values_list('materiaasignada_id', flat=True)
#             for materiaasig in alumnos.exclude(id__in=asignados):
#                 if asignados.count() >= cupomaximo:
#                     break
#                 if asignados.count() < cupomaximo:
#                     print(f"{materiaasig.matricula} -- {materiaasig.materia.asignatura} -- {grupoprofesor}")
#                     # grupoprofesor = GruposProfesorMateria.objects.get(pk=horario[0][3])
#                     if AlumnosPracticaMateria.objects.values('id').filter(materiaasignada=materiaasig, status=True).exists():
#                         alumnopractica = AlumnosPracticaMateria.objects.filter(materiaasignada=materiaasig, status=True)[0]
#                         # alumnopractica.profesormateria = grupoprofesor.profesormateria
#                         # alumnopractica.grupoprofesor = grupoprofesor
#                     else:
#                         alumnopractica = AlumnosPracticaMateria(profesormateria=grupoprofesor.profesormateria, materiaasignada=materiaasig, grupoprofesor=grupoprofesor)
#                     alumnopractica.save()
#                     if AlumnosPracticaMateria.objects.filter(status=True, grupoprofesor=grupoprofesor).values_list('materiaasignada_id', flat=True).count() >= cupomaximo:
#                         break





def proceso_asignacion_practicas_V3():
    from sga.models import Periodo, Malla, AsignaturaMalla, Carrera, Coordinacion, Notificacion, ProfesorMateria, MateriaAsignada,TipoProfesor,GruposProfesorMateria,AlumnosPracticaMateria, Notificacion
    from settings import DEBUG
    from django.db import transaction
    import math
    from sga.funciones import variable_valor
    with transaction.atomic():
        try:

                periodo_id = 336
                #, matricula__nivelmalla__orden__gt = 1
                materias_asignadas = MateriaAsignada.objects.filter(status = True, matricula__nivel__periodo_id = periodo_id, materia__practicas = True, materia__profesormateria__tipoprofesor_id__in = [2,13]).distinct('materia_id')
                for materiaasig in materias_asignadas:
                    if materiaasig.materia.asignaturamalla.malla.carrera_id.__str__() in variable_valor('LISTA_CARRERA_PARA_MATRICULA_GRUPO_PRACTICA'):
                        materiaespractica = materiaasig.materia.asignaturamalla.practicas
                        novalidar_horario_cupo_materiavirtual = materiaasig.materia.tipomateria == 2 and not variable_valor('VALIDAR_HORARIO_CUPO_MATERIA_VIRTUAL')
                        horario = []
                        profesoresmateria = None
                        if materiaespractica:
                            tipos = ProfesorMateria.objects.values_list("tipoprofesor_id", flat=True).filter(materia=materiaasig.materia, status=True, activo=True)
                            tiposprofesorpractica = TipoProfesor.objects.filter(pk__in=tipos)
                            profesoresmateria = materiaasig.materia.profesores_materia_segun_tipoprofesor_pm(tiposprofesorpractica, multiple=True)
                            for profesormateria in profesoresmateria:
                                cant_grupo = 0
                                cant_estudiantes = 0
                                cant_cupos = 0
                                bandera = False
                                grupoprofesormat = profesormateria.grupoprofesormateria()
                                if grupoprofesormat:
                                    registrados = []
                                    cant_grupo = grupoprofesormat.count()
                                    for grupo in grupoprofesormat:
                                        horario = []
                                        horarios = grupo.clases_informacion()
                                        if (horarios and grupo.cuposdisponiblesgrupoprofesor() > 0) or (novalidar_horario_cupo_materiavirtual):
                                            horario.append(["<br>".join(x for x in horarios),
                                                            profesormateria.id,
                                                            profesormateria.materia.cupo if novalidar_horario_cupo_materiavirtual else grupo.cuposdisponiblesgrupoprofesor(),
                                                            grupo.id, grupo.get_paralelopractica_display(),
                                                            grupo.profesormateria.profesor.persona.nombre_completo_inverso().__str__()])
                                            estudiantes = materiaasig.materia.asignados_a_esta_materia().exclude(pk__in = registrados)
                                            ## CUPOS
                                            gruposprofesor = GruposProfesorMateria.objects.filter(pk=horario[0][3])
                                            if bandera == False:
                                                cant_cupos = math.ceil(estudiantes.count() / cant_grupo)
                                                bandera = True
                                            for grup in gruposprofesor:
                                                grup.cupo = cant_cupos
                                                grup.save()

                                            for materiaasig in estudiantes:
                                                for grupoprofesor in gruposprofesor:
                                                    if grupoprofesor.cuposdisponiblesgrupoprofesor_congrupos() > 0:
                                                        grupoprofesor = GruposProfesorMateria.objects.get(pk=horario[0][3])
                                                        if AlumnosPracticaMateria.objects.values('id').filter(
                                                                materiaasignada=materiaasig, status=True).exists():
                                                            alumnopractica = AlumnosPracticaMateria.objects.filter(
                                                                materiaasignada=materiaasig, status=True)[0]
                                                            alumnopractica.profesormateria = grupoprofesor.profesormateria
                                                            alumnopractica.grupoprofesor = grupoprofesor
                                                        else:
                                                            alumnopractica = AlumnosPracticaMateria(
                                                                profesormateria=grupoprofesor.profesormateria,
                                                                materiaasignada=materiaasig,
                                                                grupoprofesor=grupoprofesor)
                                                        registrados.append(materiaasig.id)
                                                        alumnopractica.save()
                if DEBUG: print('PROCESO TERMINADO')
        except Exception as ex:
            transaction.set_rollback(True)
            noti = Notificacion(titulo='Error',
                                cuerpo='Ha ocurrido un error {} - Error en la linea {}'.format(
                                    ex, sys.exc_info()[-1].tb_lineno),
                                destinatario_id=38488, url="",
                                prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                                tipo=2, en_proceso=False, error=True)
            noti.save()




def reporte_datos_estudiantes_ingles():
    import xlsxwriter
    import os
    import sys
    from django.db import transaction
    from settings import SITE_STORAGE
    from datetime import datetime, timedelta
    from sga.models import Periodo, Notificacion, Matricula, MateriaAsignada, Malla, Materia, RecordAcademico, Persona
    per = Persona.objects.get(cedula='0604551580')

    try:

        dominio_sistema = 'https://sga.unemi.edu.ec'
        # output_folder = os.path.join(SITE_STORAGE, 'media')
        # if not os.path.exists(output_folder):
        #     os.makedirs(output_folder)
        #
        # nombrearchivo = f"reporte_notas_estudiantes_ingles_comp.xlsx"
        # workbook = xlsxwriter.Workbook(os.path.join(output_folder, nombrearchivo))
        # ws = workbook.add_worksheet("Estudiantes")
        #
        # # Formatos
        # fcabeceracolumna = workbook.add_format({
        #     'font_name': 'Verdana', 'bold': True, 'font_size': 10,
        #     'bg_color': '#D3D3D3', 'border': 1, 'align': 'center', 'valign': 'vcenter'
        # })
        # fceldageneral = workbook.add_format({'font_name': 'Verdana', 'font_size': 10, 'border': 1})
        # ftitulo2izq = workbook.add_format({
        #     'font_name': 'Verdana', 'bold': True, 'font_size': 14,
        #     'align': 'center', 'valign': 'vcenter'
        # })
        # fhyperlink = workbook.add_format({
        #     'font_name': 'Verdana', 'font_size': 10, 'font_color': 'blue', 'underline': 1, 'border': 1
        # })
        # fcomparaigual = workbook.add_format({
        #     'font_name': 'Verdana', 'font_size': 10, 'border': 1, 'bg_color': '#C6EFCE'
        # })
        # fcomparadistinto = workbook.add_format({
        #     'font_name': 'Verdana', 'font_size': 10, 'border': 1, 'bg_color': '#FFC7CE'
        # })
        #
        # # Títulos
        # ws.merge_range(0, 0, 0, 14, 'UNIVERSIDAD ESTATAL DE MILAGRO', ftitulo2izq)
        # ws.merge_range(1, 0, 1, 14, 'REPORTE DE ESTUDIANTES DE INGLES', ftitulo2izq)
        # ws.merge_range(2, 0, 2, 14, 'FECHA: ' + datetime.now().strftime("%d-%m-%Y"), ftitulo2izq)
        #
        # # Definir columnas
        # columnas = [
        #     (u"CARRERA", 60),
        #     (u"MODALIDAD", 20),
        #     (u"SEMESTRE", 20),
        #     (u"IDENTIFICACIÓN", 15),
        #     (u"ESTUDIANTE", 20),
        #     (u"ASIGNATURA", 50),
        #     (u"NIVEL", 20),
        #     (u"PARALELO", 20),
        #     (u"SESIÓN", 15),
        #     (u"PPL", 15),
        #     (u"EMAIL", 15),
        #     (u"EMAIL_INSTITUCIONAL", 15),
        #     (u"NOTA", 10),
        #     (u"COMPARACIÓN RECORD ACADÉMICO", 30)
        # ]
        #
        # row_num = 3
        # for col_num in range(len(columnas)):
        #     ws.write(row_num, col_num, columnas[col_num][0], fcabeceracolumna)
        #     ws.set_column(col_num, col_num, columnas[col_num][1])
        #
        # fila = 4

        periodo = Periodo.objects.get(pk=336)

        malla_ingles = Malla.objects.get(pk=353)
        id_asignaturas_ingles = malla_ingles.lista_materias_malla().filter(
            nivelmalla__orden__in=[6, 7, 5, 4, 3, 2, 1, 8, 9, 10]
        ).values_list('asignatura__id', flat=True)
        materias = Materia.objects.filter(
            status=True,
            nivel__periodo=periodo, nivel_id=1783,
            asignaturamalla__asignatura__id__in=id_asignaturas_ingles
        ).values_list('id', flat=True)

        lista_materiaasignada = MateriaAsignada.objects.filter(
            matricula__status=True,
            status=True,
            matricula__bloqueomatricula=False,
            matricula__retiradomatricula=False,
            materia__id__in=materias,
            retiramateria=False, matriculas__gte=2
        ).values_list('matricula_id', flat=True).distinct()
        print(len(lista_materiaasignada))
        c = 0
        total = 0
        for e in Matricula.objects.filter(status=True, id__in=lista_materiaasignada):
            c+=1
            print(c)
            # registro_academico = RecordAcademico.objects.filter(
            #     inscripcion=e.matricula.inscripcion,
            #     asignatura_id=e.materia.asignatura.id
            #
            # ).first()
            materias_matricula = MateriaAsignada.objects.filter(
                matricula__status=True,
                status=True, matricula=e,
                matricula__bloqueomatricula=False,
                matricula__retiradomatricula=False,
                materia__id__in=materias,
                retiramateria=False, matriculas__gte=2
            ).values_list('id', flat=True).distinct().count()

            rubro = Rubro.objects.filter(persona=e.inscripcion.persona, observacion="INGLÉS %s" % periodo.nombre).count()
            if rubro < materias_matricula:
                # estado = 'CANCELADO' if rubro.cancelado else 'PENDIENTE PAGO'
                for r in range(0, materias_matricula - rubro):
                    print(f'RUBRO {r} {e}')
                    materiaasignada = MateriaAsignada.objects.filter(
                        matricula__status=True,
                        status=True, matricula=e,
                        matricula__bloqueomatricula=False,
                        matricula__retiradomatricula=False,
                        materia__id__in=materias,
                        retiramateria=False, matriculas__gte=2).distinct().first()
                    e.nuevo_calculo_matricula_ingles(materiaasignada)
                    total +=1
            # elif e.inscripcion.persona.tiene_otro_titulo(e.inscripcion):
            #     print('RUBRO POR OTRO TITULO')
        #     if not registro_academico:
        #         continue
        #     if not registro_academico.aprobada:
        #         continue
        #     estudiante = e.matricula.inscripcion.persona
        #     _nivelsemestre = e.matricula.nivelmalla
        #     _nivel = e.materia.asignaturamalla.nivelmalla
        #     # sexo = estudiante.sexo.nombre if estudiante.sexo else "NO REGISTRA"
        #     # esppl = 'SI' if estudiante.ppl else 'NO'
        #     # email = estudiante.email
        #     # emailinst = estudiante.emailinst
        #
        #     # Comparar con record académico
        #
        #
        #
        #     if registro_academico and registro_academico.aprobada:
        #         mensaje = f"Se elimina matricula de inglés del alumno: {e} "# <-- Solo si existe registro_academico con nota 0
        #         if rubro := Rubro.objects.filter(persona=estudiante, cancelado=False, observacion="INGLÉS %s" %  periodo.nombre).first():
        #             estado = 'CANCELADO' if rubro.cancelado else 'PENDIENTE PAGO'
        #             if rubro.tiene_pagos():
        #                 mensaje+=f'-- Se mantiene el rubro: {rubro.observacion} de valor {rubro.valor} con saldo {rubro.saldo} en estado: {estado}'
        #                 # raise NameError(f"Lo sentimos, no puede desmatricularse de este módulo porque tiene pagos realizados")
        #             else:
        #                 # rubro.delete()
        #                 mensaje += f'-- Se desactiva el rubro: {rubro.observacion} de valor {rubro.valor} con saldo {rubro.saldo} en estado: {estado} por eliminación de matricula inglés'
        #                 rubro.status=False
        #                 rubro.save()
        #             from matricula.models import DetalleRubroMatricula
        #             if detallerubros := DetalleRubroMatricula.objects.filter(status=True, matricula=e.matricula, materia=e.materia).first():
        #                 detallerubros.delete()
        #         e.delete()
        #         if matriculaactual := Matricula.objects.filter(nivel__periodo_id=e.materia.nivel.periodo_id, inscripcion=e.matricula.inscripcion).exclude(retiradomatricula=True).first():
        #             if not matriculaactual.materiaasignada_set.filter(status=True):
        #                 matriculaactual.delete()
        #         user = User.objects.get(id=1)
        #         log(f'{mensaje}', None, "del", user)
        #         # break
        #         # if registro_academico and registro_academico.nota == e.valor_nombre_campo('nota'):
        #         #     comparacion = " Iguales "
        #         # else:
        #         #     comparacion = " Distintos "
        #
        #         # ws.write(fila, 0, str(e.matricula.inscripcion.carrera), fceldageneral)
        #         # ws.write(fila, 1, str(e.matricula.inscripcion.carrera.get_modalidad_display()), fceldageneral)
        #         # ws.write(fila, 2, str(_nivelsemestre), fceldageneral)
        #         # ws.write(fila, 3, str(estudiante.identificacion()), fceldageneral)
        #         # ws.write(fila, 4, str(estudiante.nombre_completo()), fceldageneral)
        #         # ws.write(fila, 5, str(e.materia.asignatura.nombre), fceldageneral)
        #         # ws.write(fila, 6, str(_nivel), fceldageneral)
        #         # ws.write(fila, 7, str(e.materia.paralelo), fceldageneral)
        #         # ws.write(fila, 8, str(e.matricula.nivel.sesion.nombre), fceldageneral)
        #         # ws.write(fila, 9, str(esppl), fceldageneral)
        #         # ws.write(fila, 10, str(email), fceldageneral)
        #         # ws.write(fila, 11, str(emailinst), fceldageneral)
        #         #
        #         # celda = 12
        #         # # for campo in e.materia.modeloevaluativo.campos():
        #         # #     nota = round(e.valor_nombre_campo(campo.nombre), campo.decimales)
        #         # #     ws.write(fila, celda, str(nota), fceldageneral)
        #         # #     celda += 1
        #         # #
        #         # # # Escribir comparación
        #         # # if comparacion == "Iguales":
        #         # #     ws.write(fila, 13, comparacion, fcomparaigual)
        #         # # else:
        #         # if registro_academico:
        #         #     estado = 'APROBADO' if registro_academico.aprobada else 'REPROBADO'
        #         #     ws.write(fila, 13, f'Asignatura {registro_academico.asignatura} - Estado {estado}', fcomparadistinto)
        #         # else:
        #         #     ws.write(fila, 13, f'No consta en record', fcomparadistinto)
        #
        # #         fila += 1
        # #
        # # workbook.close()
        #
        # noti = Notificacion(titulo='Proceso terminado',
        #                     cuerpo=f"",
        #                     destinatario=per, url="",
        #                     prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
        #                     tipo=2, en_proceso=False, error=True)
        # noti.save()

    except Exception as ex:
        err = 'Error on line {}'.format(sys.exc_info()[-1].tb_lineno)
        msg_err = f"Ocurrio un error: {ex.__str__()}. {err}"
        print('%s' % msg_err)
        noti = Notificacion(titulo='Error',
                            cuerpo='Ha ocurrido un error {} - Error en la linea {}'.format(
                                ex, sys.exc_info()[-1].tb_lineno),
                            destinatario=per, url="",
                            prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                            tipo=2, en_proceso=False, error=True)
        noti.save()


# reporte_datos_estudiantes_ingles()


duplicados = []

def reporte_datos_estudiantes_ingles_v2():
    import xlsxwriter
    import os
    import sys
    from django.db import transaction
    from settings import SITE_STORAGE
    from datetime import datetime, timedelta
    from sga.models import Periodo, Notificacion, Matricula, MateriaAsignada, Malla, Materia, RecordAcademico, Persona
    per = Persona.objects.get(cedula='0604551580')

    try:

        dominio_sistema = 'https://sga.unemi.edu.ec'
        output_folder = os.path.join(SITE_STORAGE, 'media')
        if not os.path.exists(output_folder):
            os.makedirs(output_folder)

        nombrearchivo = f"reporte_notas_estudiantes_ingles_comp.xlsx"
        workbook = xlsxwriter.Workbook(os.path.join(output_folder, nombrearchivo))
        ws = workbook.add_worksheet("Estudiantes")

        # Formatos
        fcabeceracolumna = workbook.add_format({
            'font_name': 'Verdana', 'bold': True, 'font_size': 10,
            'bg_color': '#D3D3D3', 'border': 1, 'align': 'center', 'valign': 'vcenter'
        })
        fceldageneral = workbook.add_format({'font_name': 'Verdana', 'font_size': 10, 'border': 1})
        ftitulo2izq = workbook.add_format({
            'font_name': 'Verdana', 'bold': True, 'font_size': 14,
            'align': 'center', 'valign': 'vcenter'
        })
        fhyperlink = workbook.add_format({
            'font_name': 'Verdana', 'font_size': 10, 'font_color': 'blue', 'underline': 1, 'border': 1
        })
        fcomparaigual = workbook.add_format({
            'font_name': 'Verdana', 'font_size': 10, 'border': 1, 'bg_color': '#C6EFCE'
        })
        fcomparadistinto = workbook.add_format({
            'font_name': 'Verdana', 'font_size': 10, 'border': 1, 'bg_color': '#FFC7CE'
        })

        # Títulos
        ws.merge_range(0, 0, 0, 14, 'UNIVERSIDAD ESTATAL DE MILAGRO', ftitulo2izq)
        ws.merge_range(1, 0, 1, 14, 'REPORTE DE ESTUDIANTES DE INGLES', ftitulo2izq)
        ws.merge_range(2, 0, 2, 14, 'FECHA: ' + datetime.now().strftime("%d-%m-%Y"), ftitulo2izq)

        # Definir columnas
        columnas = [
            (u"CARRERA", 60),
            (u"MODALIDAD", 20),
            (u"SEMESTRE", 20),
            (u"IDENTIFICACIÓN", 15),
            (u"ESTUDIANTE", 20),
            (u"ASIGNATURA", 50),
            (u"NIVEL", 20),
            (u"PARALELO", 20),
            (u"SESIÓN", 15),
            (u"PPL", 15),
            (u"EMAIL", 15),
            (u"EMAIL_INSTITUCIONAL", 15),
            (u"NOTA", 10),
            (u"COMPARACIÓN RECORD ACADÉMICO", 30)
        ]

        row_num = 3
        for col_num in range(len(columnas)):
            ws.write(row_num, col_num, columnas[col_num][0], fcabeceracolumna)
            ws.set_column(col_num, col_num, columnas[col_num][1])

        fila = 4

        periodo = Periodo.objects.get(pk=336)

        malla_ingles = Malla.objects.get(pk=353)
        id_asignaturas_ingles = malla_ingles.lista_materias_malla().filter(
            nivelmalla__orden__in=[6, 7, 5, 4, 3, 2, 1, 8, 9, 10]
        ).values_list('asignatura__id', flat=True)
        materias = Materia.objects.filter(
            status=True,
            nivel__periodo=periodo, nivel_id=1783,
            asignaturamalla__asignatura__id__in=id_asignaturas_ingles
        ).values_list('id', flat=True)

        lista_materiaasignada = MateriaAsignada.objects.filter(
            matricula__status=True,
            status=True,
            matricula__bloqueomatricula=False,
            matricula__retiradomatricula=False,
            materia__id__in=materias,
            retiramateria=False
        )
        # ids = HistoricoRecordAcademico.objects.filter(inscripcion_id__in=
        #                                      lista_materiaasignada.values_list('matricula__inscripcion_id', flat=True).distinct(),
        #                                      asignatura_id__in=id_asignaturas_ingles, aprobada=True, fecha__gte='2024-8-16'
        #                                      ).values_list('inscripcion_id', flat=True).distinct()

        final = lista_materiaasignada.filter(matricula_id__in=duplicados)
        print(len(final))
        c = len(final)
        for e in final:
            c-=1
            print(c)
            estudiante = e.matricula.inscripcion.persona
            # _nivelsemestre = e.matricula.nivelmalla
            # _nivel = e.materia.asignaturamalla.nivelmalla
            # sexo = estudiante.sexo.nombre if estudiante.sexo else "NO REGISTRA"
            # esppl = 'SI' if estudiante.ppl else 'NO'
            # email = estudiante.email
            # emailinst = estudiante.emailinst

            # Comparar con record académico
            registro_academico = RecordAcademico.objects.filter(
                inscripcion=e.matricula.inscripcion,
                asignatura_id=e.materia.asignatura.id

            ).first()
            if not registro_academico:
                continue
            if not registro_academico.aprobada:
                continue
            if registro_academico and registro_academico.aprobada:
                mensaje = f"Se elimina matricula de inglés del alumno: {e} "  # <-- Solo si existe registro_academico con nota 0
                if rubro := Rubro.objects.filter(persona=estudiante, cancelado=False, observacion="INGLÉS %s" % periodo.nombre).first():
                    estado = 'CANCELADO' if rubro.cancelado else 'PENDIENTE PAGO'
                    if rubro.tiene_pagos():
                        mensaje += f'-- Se mantiene el rubro: {rubro.observacion} de valor {rubro.valor} con saldo {rubro.saldo} en estado: {estado}'
                        # raise NameError(f"Lo sentimos, no puede desmatricularse de este módulo porque tiene pagos realizados")
                    else:
                        # rubro.delete()
                        mensaje += f'-- Se desactiva el rubro: {rubro.observacion} de valor {rubro.valor} con saldo {rubro.saldo} en estado: {estado} por eliminación de matricula inglés'
                        rubro.status = False
                        rubro.save()
                    from matricula.models import DetalleRubroMatricula
                    if detallerubros := DetalleRubroMatricula.objects.filter(status=True, matricula=e.matricula, materia=e.materia).first():
                        detallerubros.delete()
                e.delete()
                if matriculaactual := Matricula.objects.filter(nivel__periodo_id=e.materia.nivel.periodo_id, inscripcion=e.matricula.inscripcion).exclude(retiradomatricula=True).first():
                    if not matriculaactual.materiaasignada_set.filter(status=True):
                        matriculaactual.delete()
                user = User.objects.get(id=1)
                log(f'{mensaje}', None, "del", user)

        workbook.close()

        noti = Notificacion(titulo='Proceso terminado',
                            cuerpo=f"Archivo guardado en: {os.path.join(output_folder, nombrearchivo)}",
                            destinatario=per, url="",
                            prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                            tipo=2, en_proceso=False, error=True)
        noti.save()

    except Exception as ex:
        print(ex)
        transaction.set_rollback(True)




# from bd.models import InventarioOpcionSistema
# from laboratorio.models import LaboratorioOpcionSistema, ProcesoOpcionSistema
#
# for pregunta in InventarioOpcionSistema.objects.filter(status=True, modulo_id=48).exclude(preguntauxplora=''):
#     procesoold = pregunta.proceso
#     proceso = ProcesoOpcionSistema.objects.filter(status=True, descripcion=procesoold.descripcion).first()
#     if not proceso:
#         proceso = ProcesoOpcionSistema(descripcion=procesoold.descripcion)
#         proceso.save()
#
#     pregu = LaboratorioOpcionSistema(modulo=pregunta.modulo, nombre=pregunta.nombre, url=pregunta.url, descripcion=pregunta.descripcion,
#                                      imagen=pregunta.archivo, proceso=proceso, pregunta=pregunta.preguntauxplora, activo=True)
#     pregu.save()
#
#
#
#
# from django.db import transaction
# from settings import RUBRO_ARANCEL
# from datetime import datetime
# from decimal import Decimal
# # from idioma.models import Grupo, GrupoInscripcion, Periodo
# from sga.models import Inscripcion, RecordAcademico, AsignaturaMalla, Malla, Notificacion,ModuloMalla
#
# periodos_filtrados = Periodo.objects.filter(
#     status=True,
#     id__in=[4, 5]
# )
# # grupos_filtrados = Grupo.objects.filter(periodo__in=periodos_filtrados)
# # grupo_inscripciones_filtradas = GrupoInscripcion.objects.filter(grupo__in=grupos_filtrados).values_list('inscripcion_id', flat=True)
# grupo_inscripciones_filtradas = MateriaAsignada.objects.filter(status=True, materia__nivel_id=1790).values_list('matricula__inscripcion_id', flat=True)
#
# asignaturas_en_malla = AsignaturaMalla.objects.filter(malla_id__in=[353, 22]).values_list('asignatura_id', flat=True)
#
# mallas_sin_valor = []
# mallas_con_valor = []
# try:
#     with transaction.atomic():
#         personas_creadas = []
#         for ins in grupo_inscripciones_filtradas:
#             inscripcion = Inscripcion.objects.get(id=ins)
#             if asignatura := RecordAcademico.objects.filter(asignatura_id__in=asignaturas_en_malla,
#                                                             inscripcion=inscripcion, aprobada=False).order_by('-fecha_creacion').first():
#                 from sga.models import Periodo
#                 from sagest.models import TipoOtroRubro, Rubro
#                 from inno.models import PeriodoMalla, DetallePeriodoMalla
#                 tiporubroarancel = TipoOtroRubro.objects.filter(pk=RUBRO_ARANCEL).first()
#                 periodoactual = Periodo.objects.get(id=336)
#                 persona = inscripcion.persona
#                 if not Rubro.objects.values("id").filter(persona=persona, tipo=tiporubroarancel, observacion="INGLÉS %s" % periodoactual.nombre).exists():
#                     if not asignatura.modulomalla_id:
#                         # creditosmalla = AsignaturaMalla.objects.filter(malla_id__in=[353, 22], asignatura=asignatura.asignatura, creditos__gt=0).first().creditos
#                         if modulo_malla := ModuloMalla.objects.filter(asignatura=asignatura.asignatura, malla=inscripcion.mi_malla(), creditos__gt=0).first():
#                             creditosmalla = modulo_malla.creditos
#                     # creditosmalla = AsignaturaMalla.objects.filter(malla_id__in=[353, 22], asignatura=asignatura.asignatura, creditos__gt=0).first().creditos
#                     else:
#                         creditosmalla = ModuloMalla.objects.filter(id=asignatura.modulomalla_id, creditos__gt=0).first().creditos
#
#                     if not inscripcion.coordinacion_id in [9, 7]:
#                         cobro = 2
#                         # if not asignatura.materiaregular:
#                         #     raise NameError(f'No tiene materia regurlar {asignatura}')
#                         periodo = Periodo.objects.get(id=126)
#                         gruposocioeconomico = persona.grupoeconomico()
#
#                         valorgrupoeconomico = 0
#                         if periodo.tipocalculo == 1:
#                             periodogruposocioeconomico = gruposocioeconomico.periodogruposocioeconomico_set.filter(periodo=periodo, status=True)
#                             if not periodogruposocioeconomico.values("id").exists():
#                                 raise NameError(u"Grupo socioeconomico no tiene configurado valores de créditos")
#                             valorgrupoeconomico = periodogruposocioeconomico.first().valor
#                         else:
#                             malla = inscripcion.mi_malla()
#                             if not malla:
#                                 raise NameError(u"Malla sin configurar")
#                             periodomalla = PeriodoMalla.objects.filter(periodo=periodo, malla=malla, status=True)
#                             if not periodomalla.values("id").exists():
#                                 periodomalla = PeriodoMalla.objects.filter(periodo=periodo, malla__carrera=malla.carrera, status=True)
#
#                                 # raise NameError(f"Malla no tiene configurado valores de cobro {malla} {malla.id} {periodo}")
#                             periodomalla = periodomalla.first()
#                             detalleperiodomalla = DetallePeriodoMalla.objects.filter(periodomalla=periodomalla,
#                                                                                      gruposocioeconomico=gruposocioeconomico,
#                                                                                      status=True)
#                             if not detalleperiodomalla.values("id").exists():
#                                 raise NameError(u"Malla en grupo socioeconomico no tiene configurado valores de cobro")
#                             valorgrupoeconomico = detalleperiodomalla.first().valor
#
#                         porcentaje_gratuidad = periodo.porcentaje_gratuidad
#                         costo_materia_total = 0
#
#                         if cobro > 0:
#                             creditos_para_cobro = creditosmalla
#                             if cobro == 2:
#                                 costo_materia = Decimal(
#                                     Decimal(creditos_para_cobro).quantize(Decimal('.01')) * valorgrupoeconomico).quantize(
#                                     Decimal('.01'))
#                             else:
#                                 costo_materia = Decimal(
#                                     Decimal(creditos_para_cobro).quantize(Decimal('.01')) * valorgrupoeconomico).quantize(
#                                     Decimal('.01'))
#                             costo_materia_total += costo_materia
#
#                         if costo_materia_total > 0:
#                             # raise NameError(f'{persona} -- Periodo materia: {periodoactual.nombre} -- valor rubro: {costo_materia_total} -- Creditos: {creditos_para_cobro} -- valorgrupo: {valorgrupoeconomico}')
#                             if not Rubro.objects.values("id").filter(persona=persona, tipo=tiporubroarancel, observacion="INGLÉS %s" % periodoactual.nombre).exists():
#                                 rubro = Rubro(tipo=tiporubroarancel,
#                                               persona=persona,
#                                               nombre=tiporubroarancel.nombre + ' - ' + periodoactual.nombre,
#                                               cuota=1,
#                                               fecha=datetime.now().date(),
#                                               fechavence=datetime(2024, 9, 13).date(),
#                                               valor=costo_materia_total,
#                                               iva_id=1,
#                                               valoriva=0,
#                                               valortotal=costo_materia_total,
#                                               saldo=costo_materia_total,
#                                               cancelado=False,
#                                               observacion="INGLÉS %s" % periodoactual.nombre)
#                                 rubro.save()
#                                 personas_creadas.append(persona.id)
#         noti = Notificacion(cuerpo=f'Creados: {personas_creadas}',
#                             titulo=f'Creados: {len(personas_creadas)}',
#                             destinatario_id=33490,
#                             prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now(),
#                             tipo=2, en_proceso=False)
#         noti.save()
# except Exception as ex:
#     print(ex)
#     noti = Notificacion(cuerpo=f'Error: {ex}',
#                         titulo=f'Error: {ex}',
#                         destinatario_id=33490,
#                         prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now(),
#                         tipo=2, en_proceso=False)
#     noti.save()
#
#

causas = [
[1,
     [1,  'Nunca realizó la matrícula del cupo asignado (No efectivizó el cupo) o no posee información del aspirante']
 ],
[2,
     [6,  'Actualmente es estudiante de la institución en nivelación. Se encuentra asistiendo a clases.'],
     [7,  'Se matriculó y posterior se retiró, anuló, desertó, dejó de asistir o perdió la primera matrícula del curso de nivelación de carrera. (La institución no puede recibirlo).'],
     [8,  'Se matriculó y posterior se retiró, anuló, desertó, dejó de asistir o perdió la primera matrícula del curso de nivelación de carrera. (La institución si puede recibirlo).'],
     [26, 'Reprobó la nivelación en segunda matricula.'],
     [9,  'Aprobó la nivelación de carrera, no se matriculó en primer semestre. La institución no puede recibirlo.'],
     [10, 'Aprobó la nivelación de carrera, no se matriculó en primer semestre. Puede matricularse en primer semestre.'],
     [29, 'Anuló la matrícula. La anulación se da por contravenir la Ley o la normativa aplicable (Art. 72 Reglamento Régimen Académico).']
 ],
[3,
     [13, 'El estudiante registra matrícula en carrera.'],
     [14, 'Agotó el número máximo de matrículas permitidas en la carrera (Impedimento académico).']
 ],
[4,
     [24, 'Graduado - Egresado']
 ]
]

def es_graduado(inscripcion):
    return inscripcion.graduado_set.values("id").filter(status=True, estadograduado=True).exists()

def es_egresado(inscripcion):
    return inscripcion.egresado_set.values("id").exists()

def es_matriculado_periodo(inscripcion):
    return inscripcion.matricula_set.filter(nivel__periodo_id=336, retiradomatricula=False).exists()

def tiene_mas_tres_matricula(carrera, inscripcion):
    pass

def reprobo_nivelacion_segunda_vez(carrera, inscripcion):
    return inscripcion.matricula_set.filter(nivel__periodo_id=336, retiradomatricula=False).exists()

def esta_en_nivelacion(carrera, inscripcion):
    pass

import openpyxl
print('Inicio del proceso matriz')
miarchivo = openpyxl.load_workbook("matriz_seguimiento_2012_2024.xlsx")
lista = miarchivo.active
totallista = lista.rows
a = 0
for fila in totallista:
    a+=1
    if a<=1:
        continue
    if fila[12].value in [None, '', "", " ", ' ', 'null']:
        continue
    cedula = fila[8].value
    if cedula in [None, '', "", " ", ' ', 'null']:
        break
    carreras = [int(i) for i in fila[12].value.split(',')] #carreras de grado
    carreras_n = fila[13].value #carreras de nivelacion
    inscripcion = Inscripcion.objects.filter(persona__cedula=cedula, carrera_id__in=carreras)
    if not inscripcion:
        continue
    if inscripcion.count() > 1:
        for i in inscripcion:
            if es_graduado(i) or  es_egresado(i):
                fila[17].value = causas[3][1][1]
                break
    else:
        if es_graduado(inscripcion.first()) or es_egresado(inscripcion.first()):
            fila[11].value = causas[3][1][1]
            continue
miarchivo.save('matriz_seguimiento_2012_2024.xlsx')


