# -*- coding: UTF-8 -*-
import json
import random

from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Q
from django.http import HttpResponseRedirect, JsonResponse, HttpResponse
from django.shortcuts import render

from openpyxl import workbook as openxl
from openpyxl.styles import Font as openxlFont
from openpyxl.styles.alignment import Alignment as alin

from decorators import secure_module
from sagest.commonviews import anio_ejercicio
from sagest.forms import TipoOtroRubroForm, TipoOtroRubroPartidaForm
from sagest.models import  TipoOtroRubro, TipoOtroRubroSaldoPartida, PartidasSaldo, AnioEjercicio
from sga.commonviews import adduserdata
from sga.funciones import log, MiPaginador


@login_required(redirect_field_name='ret', login_url='/loginsagest')
@secure_module
@transaction.atomic()
def view(request):
    data = {}
    adduserdata(request, data)
    persona = request.session['persona']
    if 'aniofiscalpresupuesto' in request.session:
        anio = request.session['aniofiscalpresupuesto']
    else:
        anio = anio_ejercicio().anioejercicio
    usuario = request.user
    if request.method == 'POST':
        action = request.POST['action']

        if action == 'addrubros':
            try:
                form = TipoOtroRubroForm(request.POST)
                if form.is_valid():
                    registro = TipoOtroRubro(nombre=form.cleaned_data['nombre'],
                                             partida=form.cleaned_data['partida'],
                                             unidad_organizacional=form.cleaned_data['unidad_organizacional'],
                                             programa=form.cleaned_data['programa'],
                                             interface=form.cleaned_data['interface'],
                                             valor=form.cleaned_data['valor'],
                                             ivaaplicado=form.cleaned_data['ivaaplicado'],
                                             activo=form.cleaned_data['activo'],
                                             tiporubro=form.cleaned_data['tipo'],
                                             exportabanco=form.cleaned_data['exportabanco'],
                                             nofactura=form.cleaned_data['nofactura'])
                    registro.save(request)
                    log(u'Registro nuevo rubro: %s' % registro, request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                     raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": "Nombre de rubro ya existe."})

        if action == 'asignar':
            try:
                form = TipoOtroRubroPartidaForm(request.POST)
                if form.is_valid():
                    rubro = TipoOtroRubro.objects.get(pk=int(request.POST['id']))
                    rubropartida = TipoOtroRubroSaldoPartida(tipootrorubro=rubro,
                                                             partidassaldo=form.cleaned_data['partida'])
                    rubropartida.save(request)
                    log(u'Registro asigno partida: %s' % rubropartida, request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                     raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": "Error al guardar los datos."})

        if action == 'editasignar':
            try:
                form = TipoOtroRubroPartidaForm(request.POST)
                if form.is_valid():
                    rubropartida = TipoOtroRubroSaldoPartida.objects.get(pk=int(request.POST['id']))
                    partida = rubropartida.partidassaldo
                    if form.cleaned_data['partida']:
                        partida = form.cleaned_data['partida']
                    rubropartida.partidassaldo=partida
                    rubropartida.save(request)
                    log(u'Registro asigno partida: %s' % rubropartida, request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                     raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": "Error al guardar los datos."})

        if action == 'editrubros':
            try:
                form = TipoOtroRubroForm(request.POST)
                if form.is_valid():
                    registro = TipoOtroRubro.objects.get(pk=int(request.POST['id']))
                    partidaanterior = registro.partida
                    if not registro.en_uso(anio):
                        registro.partida = form.cleaned_data['partida']
                    registro.nombre = form.cleaned_data['nombre']
                    registro.unidad_organizacional = form.cleaned_data['unidad_organizacional']
                    registro.programa = form.cleaned_data['programa']
                    registro.valor = form.cleaned_data['valor']
                    registro.interface = form.cleaned_data['interface']
                    registro.ivaaplicado = form.cleaned_data['ivaaplicado']
                    registro.activo = form.cleaned_data['activo']
                    registro.nofactura = form.cleaned_data['nofactura']
                    registro.tiporubro = form.cleaned_data['tipo']
                    registro.exportabanco = form.cleaned_data['exportabanco']
                    registro.save(request)
                    if registro.partida != partidaanterior:
                        detalle = registro.partida_saldo(anio)
                        if detalle:
                            detalle.delete()
                    log(u'Registro modificado Rubros: %s' % registro, request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                     raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": "Error al editar los datos."})

        if action == 'deleterubros':
            try:
                campo = TipoOtroRubro.objects.get(pk=request.POST['id'], status=True)
                if campo.en_uso(anio):
                    return JsonResponse({"result": "bad", "mensaje": "El campo se encuentra en uso."})
                campo.delete()
                log(u'Elimino campos contratos: %s' % campo, request, "del")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": "Error al eliminar los datos."})

        return JsonResponse({"result": "bad", "mensaje": "Solicitud Incorrecta."})
    else:
        if 'action' in request.GET:
            action = request.GET['action']

            if action == 'addrubros':
                try:
                    data['title'] = u'Nuevo Rubro'
                    data['form'] = TipoOtroRubroForm()
                    return render(request, "pre_rubros/addrubros.html", data)
                except Exception as ex:
                    pass

            if action == 'editrubros':
                try:
                    data['title'] = u'Modificación Rubro'
                    data['tipootrorubro'] = tipootrorubro = TipoOtroRubro.objects.get(pk=request.GET['id'])
                    form = TipoOtroRubroForm(initial={'nombre': tipootrorubro.nombre,
                                                      'partida': tipootrorubro.partida,
                                                      'unidad_organizacional': tipootrorubro.unidad_organizacional,
                                                      'programa': tipootrorubro.programa,
                                                      'interface': tipootrorubro.interface,
                                                      'ivaaplicado': tipootrorubro.ivaaplicado,
                                                      'valor': tipootrorubro.valor,
                                                      'activo': tipootrorubro.activo,
                                                      'tipo': tipootrorubro.tiporubro,
                                                      'exportabanco': tipootrorubro.exportabanco,
                                                      'nofactura': tipootrorubro.nofactura})
                    form.edit(tipootrorubro, anio)
                    data['form'] = form
                    return render(request, "pre_rubros/editrubros.html", data)
                except Exception as ex:
                    pass

            if action == 'asignar':
                try:
                    data['title'] = u'Asignar Partida Saldo'
                    data['tipootrorubro'] = tipootrorubro = TipoOtroRubro.objects.get(pk=request.GET['id'])
                    form = TipoOtroRubroPartidaForm()
                    form.adicionar(anio, tipootrorubro.partida)
                    data['form'] = form
                    return render(request, "pre_rubros/asignar.html", data)
                except Exception as ex:
                    pass

            if action == 'editasignar':
                try:
                    data['title'] = u'Asignar Partida Saldo'
                    data['tipootrorubro'] = tipootrorubro = TipoOtroRubro.objects.get(pk=request.GET['id'])
                    data['rubropartida'] = rubropartida = tipootrorubro.tipootrorubrosaldopartida_set.all()[0]
                    form = TipoOtroRubroPartidaForm(initial={'partida': rubropartida.partidassaldo})
                    form.edit(rubropartida, anio, tipootrorubro.partida)
                    data['form'] = form
                    return render(request, "pre_rubros/editasignar.html", data)
                except Exception as ex:
                    pass

            if action == 'deleterubros':
                try:
                    data['title'] = u'Eliminar Rubro'
                    data['rubro'] = TipoOtroRubro.objects.get(pk=request.GET['id'])
                    return render(request, 'pre_rubros/delete.html', data)
                except Exception as ex:
                    pass

            if action == 'cambioperiodo':
                try:
                    anio = AnioEjercicio.objects.get(id=int(request.GET['id']))
                    request.session['aniofiscalpresupuesto'] = anio.anioejercicio
                except Exception as ex:
                    pass

            return HttpResponseRedirect(request.path)
        else:
            data['title'] = u'Rubros'
            search = None
            tipo = None
            url_vars = ''

            if 's' in request.GET:
                if request.GET['s'] != '':
                    search = request.GET['s'].strip().upper()

            if search:
                rubro = TipoOtroRubro.objects.filter(Q(nombre__icontains=search) |
                                                     Q(partida__codigo__icontains=search) |
                                                     Q(partida__nombre__icontains=search))
                url_vars += "&s={}".format(search)
            else:
                rubro = TipoOtroRubro.objects.filter(status=True)
            paging = MiPaginador(rubro, 25)
            p = 1
            try:
                paginasesion = 1
                if 'paginador' in request.session:
                    paginasesion = int(request.session['paginador'])
                if 'page' in request.GET:
                    p = int(request.GET['page'])
                else:
                    p = paginasesion
                try:
                    page = paging.page(p)
                except:
                    p = 1
                page = paging.page(p)
            except:
                page = paging.page(p)
            request.session['paginador'] = p
            data['paging'] = paging
            data['rangospaging'] = paging.rangos_paginado(p)
            data['page'] = page
            data['search'] = search if search else ""
            data['anio'] = anio
            data['mianio'] = anio
            data['total']=len(rubro)
            data['rubros'] = page.object_list
            data['anios'] = AnioEjercicio.objects.all()
            data['url_vars'] = url_vars
            if 'exportar_excel' in request.GET:
                wb = openxl.Workbook()
                wb["Sheet"].title = "Reporte_Rubros"
                ws = wb.active
                style_title = openxlFont(name='Arial', size=16, bold=True)
                style_cab = openxlFont(name='Arial', size=10, bold=True)
                alinear = alin(horizontal="center", vertical="center")
                response = HttpResponse(content_type="application/ms-excel")
                response[
                    'Content-Disposition'] = 'attachment; filename=Reporte de rubros' + '-' + random.randint(
                    1, 10000).__str__() + '.xlsx'
                ws.column_dimensions['B'].width = 30
                ws.column_dimensions['C'].width = 30
                ws.column_dimensions['D'].width = 30
                ws.column_dimensions['E'].width = 30
                ws.column_dimensions['F'].width = 30
                ws.column_dimensions['G'].width = 15
                ws.column_dimensions['H'].width = 15
                ws.column_dimensions['I'].width = 15
                ws.column_dimensions['J'].width = 15
                ws.column_dimensions['K'].width = 15
                ws.merge_cells('A1:K1')
                ws['A1'] = 'REPORTE DE RUBROS'
                celda1 = ws['A1']
                celda1.font = style_title
                celda1.alignment = alinear

                columns = [u"N°", u"RUBRO", u"TIPO", u"PARTIDA", u"PROGRAMA", "UNIDAD. ORG",
                           u"VALOR PREDET.", u"IVA", u"INTERFACE", "EXPORTA", "ACTIVO"
                           ]
                row_num = 3
                for col_num in range(0, len(columns)):
                    celda = ws.cell(row=row_num, column=(col_num + 1), value=columns[col_num])
                    celda.font = style_cab
                mensaje='NO REGISTRA'
                row_num = 4
                numero = 0
                for list in rubro:
                    numero += 1
                    ws.cell(row=row_num, column=1, value=numero)
                    ws.cell(row=row_num, column=2, value=str(list.nombre))
                    ws.cell(row=row_num, column=3, value=list.get_tiporubro_display())
                    ws.cell(row=row_num, column=4, value=str(list.partida))
                    ws.cell(row=row_num, column=5, value=str(list.programa) if list.programa else mensaje)
                    ws.cell(row=row_num, column=6, value=str(list.unidad_organizacional)if list.unidad_organizacional else mensaje)
                    ws.cell(row=row_num, column=7, value=str(list.valor))
                    ws.cell(row=row_num, column=8, value=str(list.ivaaplicado))
                    ws.cell(row=row_num, column=9, value="SI" if list.interface else "NO")
                    ws.cell(row=row_num, column=10, value="SI" if list.exportabanco else "NO")
                    ws.cell(row=row_num, column=11, value="SI" if list.activo else "NO")
                    row_num += 1
                wb.save(response)
                return response
            return render(request, 'pre_rubros/view.html', data)