# -*- coding: latin-1 -*-
import json
import calendar
import random
from django.contrib import messages
from datetime import datetime, date, timedelta
from django.contrib.admin.models import LogEntry
from django.contrib.auth.decorators import login_required
from django.http import HttpResponseRedirect, HttpResponse, JsonResponse
from django.shortcuts import render
from django.template.context import Context
from django.template.loader import get_template
from django.db.models import Q, F, Value, Count, Case, When, ExpressionWrapper, FloatField, Subquery, OuterRef, Exists
from django.db.models.functions import Concat, Coalesce
from django.db import transaction, connections, models
from decorators import secure_module, last_access
from sga.commonviews import adduserdata, obtener_reporte, traerNotificaciones
from sga.excelbackground import limpiar_cache_encuesta_silabo
from sga.funciones import MiPaginador, log, Round2, convertir_lista, Round4, generar_nombre, variable_valor
from sga.models import Materia, ProfesorMateria, Profesor, Modulo, Clase, ModuloGrupo, Modalidad, TemaAsistencia, \
    SubTemaAsistencia, SubTemaAdicionalAsistencia, Silabo, SilaboSemanal, ResponsableCoordinacion, Carrera, Malla, \
    NivelMalla, Asignatura, Paralelo, Notificacion, Coordinacion, CoordinadorCarrera, EncuestaGrupoEstudiantes, \
    Inscripcion, Matricula, MateriaAsignada, Encuesta, ProfesorDistributivoHoras, Persona, FotoPersona
from sga.funcionesxhtml2pdf import conviert_html_to_pdf
from sga.templatetags.sga_extras import encrypt, informe_actividades_mensual_docente_v4_extra
from settings import DEBUG, SITE_STORAGE, GENERAR_TUMBAIL
from inno.models import EncuestaGrupoEstudianteSeguimientoSilabo, InscripcionEncuestaEstudianteSeguimientoSilabo, RespuestaPreguntaEncuestaSilaboGrupoEstudiantes, HistoricoCambioFotoPerfil
from inno.forms import InscripcionEncuestaEstudianteSeguimientoSilaboForm, EditFotoEstudiante
from mobile.views import make_thumb_fotopersona, make_thumb_picture
from openpyxl import workbook as openxl
from openpyxl.styles import Font as openxlFont
from openpyxl.styles.alignment import Alignment as alin


@login_required(redirect_field_name='ret', login_url='/loginsga')
@transaction.atomic()
@last_access
# @secure_module
def view(request):
    data = {}
    adduserdata(request, data)
    persona_user = request.session['persona']
    fechahoy = datetime.now().date()
    perfilprincipal = request.session['perfilprincipal']
    # if not perfilprincipal.es_administrativo():
    #     return HttpResponseRedirect("/?info=Solo los perfiles administrativos pueden ingresar al modulo.")
    if not str(persona_user.id) in (variable_valor('LISTA_ACCESO_EDIT_FOTO')):
        return HttpResponseRedirect("/?info=Solo los usuarios permitidos pueden ingresar al modulo.")

    data["acceso_reporte"] = False
    if persona_user.id in [30097, 38488]:
        data["acceso_reporte"] = True

    data['periodo'] = periodo = request.session['periodo']
    dominio_sistema = 'https://sga.unemi.edu.ec'

    if DEBUG:
        dominio_sistema = 'http://localhost:8000'

    data["DOMINIO_DEL_SISTEMA"] = dominio_sistema


    if request.method == 'POST':
        if 'action' in request.POST:
            action = request.POST['action']

            if action == 'editFoto':
                try:
                    if 'foto' in request.FILES:
                        fotofile = request.FILES['foto']
                        if fotofile.size > 8194304:
                            raise NameError(u"Archivo mayor a 8Mb.")
                        fotofileod = fotofile._name
                        ext = fotofileod[fotofileod.rfind("."):]
                        if not ext in ['.jpg', '.jpeg']:
                            raise NameError(u"Solo archivo con extensión .jpg o .jpeg")
                        persona = Persona.objects.get(pk=int(request.POST['id']))
                        form = EditFotoEstudiante(persona, request.FILES)
                        if form.is_valid():
                            fotofile._name = generar_nombre("foto_", fotofile._name)
                            foto = persona.foto()
                            if foto:
                                foto.foto = fotofile
                            else:
                                foto = FotoPersona(persona=persona,
                                                   foto=fotofile)
                            foto.save(request)
                            eMatricula = Matricula.objects.filter(status = True, inscripcion__persona__cedula__icontains=persona.cedula, nivel__periodo = periodo, retiradomatricula = False, inscripcion__coordinacion_id__in=[1,2,3,4,5])
                            for matricula in eMatricula:
                                historico_registro = HistoricoCambioFotoPerfil(
                                    cedula= matricula.inscripcion.persona.cedula,
                                    nombres = matricula.inscripcion.persona,
                                    facultad_id= matricula.inscripcion.coordinacion.id,
                                    facultad = matricula.inscripcion.coordinacion,
                                    carrera_id=matricula.inscripcion.carrera.id,
                                    carrera = matricula.inscripcion.carrera,
                                    nivel = matricula.nivelmalla,
                                    cedula_persona_registro = persona_user.cedula,
                                    persona_registro = persona_user,
                                    periodo_id = periodo.id,
                                    periodo = periodo,
                                    fecha = datetime.now().date(),
                                    foto = dominio_sistema+'/media/'+str(foto.foto)
                                    )
                                historico_registro.save(request)
                            log(u'Se actualizó la foto de: %s' % persona, request, "edit")
                            make_thumb_picture(persona)
                            if GENERAR_TUMBAIL:
                                make_thumb_fotopersona(persona)
                        else:
                            for k, v in form.errors.items():
                                raise NameError(v[0])
                        return JsonResponse({"result": True})
                    else:
                        raise NameError(u"Debe subir una foto válida")

                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": False, 'message': str(ex)})

        return JsonResponse({"result": "bad", "mensaje": u"Solicitud Incorrecta."})
    else:
        if 'action' in request.GET:
            data['action'] = action = request.GET['action']

            if action == 'editFoto':
                try:
                    data['action'] = 'editFoto'
                    data['id'] = request.GET['id']
                    form = EditFotoEstudiante()
                    data['form'] = form
                    template = get_template("adm_editfotoestudiantes/modal/cargarFoto.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return HttpResponseRedirect(request.path + f'?info={ex=}')

            elif action == 'reporteestudiantes':
                try:
                    coordinacion = int(request.GET['cod_facultad'])
                    codcarrera = int(request.GET['cod_carrera'])
                    if coordinacion <= 0:
                        return JsonResponse({"result": False, "mensaje": u"Debe seleccionar una coordinación válida"})
                    if codcarrera <= 0:
                        return JsonResponse({"result": False, "mensaje": u"Debe seleccionar una carrera válida"})

                    listado = HistoricoCambioFotoPerfil.objects.filter(status = True, carrera_id = codcarrera, facultad_id = coordinacion).order_by('nivel')
                    if listado:
                        wb = openxl.Workbook()
                        wb["Sheet"].title = "Estudiantes_actualizados"
                        ws = wb.active
                        style_title = openxlFont(name='Arial', size=16, bold=True)
                        style_cab = openxlFont(name='Arial', size=10, bold=True)
                        alinear = alin(horizontal="center", vertical="center")
                        response = HttpResponse(content_type="application/ms-excel")
                        response[
                            'Content-Disposition'] = 'attachment; filename=Reporte de estudiantes' + '-' + random.randint(
                            1, 10000).__str__() + '.xlsx'
                        ws.column_dimensions['B'].width = 20
                        ws.column_dimensions['C'].width = 20
                        ws.column_dimensions['D'].width = 20
                        ws.column_dimensions['E'].width = 20
                        ws.column_dimensions['F'].width = 20
                        ws.column_dimensions['G'].width = 20
                        ws.column_dimensions['H'].width = 20
                        ws.column_dimensions['I'].width = 20
                        ws.merge_cells('A1:I1')
                        ws['A1'] = 'LISTADO DE POSTULANTES'
                        celda1 = ws['A1']
                        celda1.font = style_title
                        celda1.alignment = alinear

                        columns = [u"N°", u"CEDULA", u"ESTUDIANTE", u"FACULTAD", u"CARRERA",
                                   u"NIVEL", u"FOTO", u"ENCARGADO", u"FECHA DE ACTUALIZACIÓN",
                                   ]
                        row_num = 2
                        for col_num in range(0, len(columns)):
                            celda = ws.cell(row=row_num, column=(col_num + 1), value=columns[col_num])
                            celda.font = style_cab
                        row_num = 3
                        numero = 1
                        historico = HistoricoCambioFotoPerfil.objects.filter(status=True)
                        for list in historico:
                            ws.cell(row=row_num, column=1, value=numero)
                            ws.cell(row=row_num, column=2, value=str(list.cedula))
                            ws.cell(row=row_num, column=3, value=str(list.nombres))
                            ws.cell(row=row_num, column=4, value=str(list.facultad))
                            ws.cell(row=row_num, column=5, value=str(list.carrera))
                            ws.cell(row=row_num, column=6, value=str(list.nivel))
                            ws.cell(row=row_num, column=7, value=str(list.foto))
                            ws.cell(row=row_num, column=8, value=str(list.persona_registro))
                            ws.cell(row=row_num, column=9, value=str(list.fecha))
                            row_num += 1
                            numero += 1
                        wb.save(response)
                        return response
                    else:
                        return JsonResponse({"result": False, "mensaje": u"No existen estudiantes en la carrera seleccionada"})

                    # noti = Notificacion(cuerpo='Generación de reporte de excel en progreso',
                    #                     titulo='Reporte consolidado de cumplimiento de actividades y recursos de aprendizaje',
                    #                     destinatario=persona_user,
                    #                     url='',
                    #                     prioridad=1, app_label='SGA',
                    #                     fecha_hora_visible=datetime.now() + timedelta(days=1), tipo=2,
                    #                     en_proceso=True)
                    # noti.save(request)
                    return JsonResponse({"result": True,
                                         "mensaje": u"El reporte se está realizando. Verifique su apartado de notificaciones después de unos minutos.",
                                         "btn_notificaciones": traerNotificaciones(request, data, persona_user)})

                    return JsonResponse({"result": True,
                                         "mensaje": u"El reporte se está realizando. Verifique su apartado de notificaciones después de unos minutos.",
                                         "btn_notificaciones": traerNotificaciones(request, data, persona)})
                except Exception as ex:
                    return HttpResponseRedirect(f"{request.path}?info={ex=}")
                    return JsonResponse({"result": False,"mensaje": u"Ha ocurrido un error en la generación del reporte."})

            elif action == 'carrerascoordinacion':
                try:
                    facultad = Coordinacion.objects.values_list('id',flat=True).filter(pk=int(request.GET['id']))
                    if int(request.GET['id']) == 2:
                        facultad = [2,3]
                    lista = []
                    idcarreras = ProfesorMateria.objects.values_list('materia__asignaturamalla__malla__carrera_id',
                                                                     flat=True).filter(tipoprofesor_id__in=[1, 14],
                                                                                       materia__nivel__periodo=periodo,
                                                                                       materia__asignaturamalla__malla__carrera__coordinacion__id__in=facultad,
                                                                                       materia__status=True,
                                                                                       status=True).exclude(
                        materia_id__in=Materia.objects.values_list('id').filter(
                            asignaturamalla__malla__carrera_id__in=[1, 3], asignaturamalla__nivelmalla_id__in=[7, 8, 9],
                            nivel__periodo=periodo, status=True))
                    carreras = Carrera.objects.filter(pk__in=idcarreras)
                    for carrera in carreras:
                        lista.append([carrera.id, carrera.__str__()])
                    return JsonResponse({'result': 'ok', 'lista': lista})
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})


            return HttpResponseRedirect(request.path)

        else:
            try:
                data['title'] = u'Edición de foto de estudiantes'
                data['ePeriodo'] = periodo
                data['eMatriculas'] = None
                paginado = 1
                if 'search' in request.GET:
                    url_vars = ''
                    search = request.GET['search'].strip()
                    url_vars += f'&s={search}'
                    ss = search.split(' ')
                    filtro = Q(status=True) & Q(nivel__periodo=periodo) & Q(inscripcion__coordinacion_id__in=[1,2,3,4,5])
                    if len(ss) == 1:
                        filtro = filtro & (Q(inscripcion__persona__cedula__icontains=search) |
                                           Q(inscripcion__persona__pasaporte__icontains=search))
                    else:
                        filtro = filtro & (Q(inscripcion__persona__cedula__icontains=search) |
                                           Q(inscripcion__persona__pasaporte__icontains=search))
                    eMatriculas = Matricula.objects.filter(filtro).order_by('inscripcion__persona__apellido1',
                                                                            'inscripcion__persona__apellido2').select_related(
                        'inscripcion')
                    data['cantMatriculas'] = cantMatriculas = eMatriculas.count()
                    if cantMatriculas > 1:
                        data['matriculas'] = eMatriculas
                    paging = MiPaginador(eMatriculas, paginado)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data['eMatriculas'] = page.object_list
                    data['url_vars'] = url_vars
                data['coordinaciones'] = Coordinacion.objects.filter(pk__in=[1, 2, 4, 5],status=True)
                return render(request, "adm_editfotoestudiantes/view.html", data)
            except Exception as ex:
                return HttpResponseRedirect(f"/?info=No puede acceder al módulo")

