# -*- coding: latin-1 -*-
from datetime import datetime
import json
import xlwt
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Sum
from django.contrib import messages
from django.forms import model_to_dict
from django.http import HttpResponseRedirect, JsonResponse, HttpResponse
from django.shortcuts import render
from django.template.loader import get_template
from django.template.context import Context
from django.db.models.query_utils import Q
from decorators import secure_module, last_access
from helpdesk.forms import HdSolicitarIncidenteFrom, HdGrupoSistemaEquipoForm, HdBienForm, HdConfFrecuenciaForm, \
    HdFrecuenciaForm, HdPlanAprobacionForm, HdCronogramaMantenimientoForm, HdDetCronogramaMantenimientoForm, \
    HdReparacionForm, HdDetalle_ReparacionForm, HdMaterialMantenimientoForm, HdMaterialMantenimiento_MaterialForm, \
    HdCronogramaMantenimientoSemForm, HdDetCronogramaMantenimientoSemForm, HdPresupuestoRecursoForm, HdMantenimientosActivosForm,\
    TipoBienForm, TareasMantenimientoForm, HdMantenimientoGruDaniosForm, HdPiezaPartesForm,ResponsableActivoTraspasoForm
from helpdesk.models import HdIncidente, \
    HdBloqueUbicacion, HdSubCategoria, HdDetalle_SubCategoria, HdDetEncuestas, HdRespuestaEncuestas, \
    HdCabRespuestaEncuestas, ActivoFijo, HdGrupo, HdCategoria, HdCausas, HdTipoIncidente, ActivosSinCodigo, \
    HdGrupoSistemaEquipo, HdBien, HdConfFrecuencia, HdFrecuencia, HdPlanAprobacion, HdCronogramaMantenimiento, \
    HdDetCronogramaMantenimiento, HdReparacion, HdDetalle_Reparacion, HdMaterialMantenimiento, \
    HdMaterialMantenimiento_Material, HdMaterialMantenimiento_Responsable, HdDetalle_Grupo, \
    HdCronogramaMantenimientoSem, HdDetCronogramaMantenimientoSem, HdPresupuestoRecurso, TIPO_MANTENIMIENTO, \
    HdDetMantenimientosActivos, HdMantenimientoGruCategoria, HdMantenimientoGruDanios, HdGruposCategoria, \
    HdTareasActivosPreventivos, HdTareasActivosPreventivosDanios, HdPiezaParteActivosPreventivos, HdPiezaPartes, SolicitudConfirmacionMantenimiento, HistorialSolicitudConfirmacionMantenimiento
from sga.commonviews import adduserdata
from sga.funciones import MiPaginador, log, generar_nombre, convertir_fecha, notificacion
from sga.funcionesxhtml2pdf import conviert_html_to_pdf
from sga.models import Persona, MONTH_CHOICES
from sagest.models import HdBloque, AnioEjercicio, ActivoFijoInventarioTecnologico, ActivoTecnologico, MantenimientoGruDanios, MantenimientoGruCategoria, ESTADO_DANIO, SolicitudActivos, \
    Proveedor, IngresoProducto
from sagest.forms import ProveedorForm
from openpyxl import Workbook
from rest_framework import reverse

from decorators import secure_module
from xlwt import *
import random
from sga.templatetags.sga_extras import encrypt
#
# @login_required(redirect_field_name='ret', login_url='/loginsga')
# @secure_module
@last_access
@transaction.atomic()
def view(request):
    global ex
    data = {}
    adduserdata(request, data)
    persona = request.session['persona']
    if request.method == 'POST':
        action = request.POST['action']
        if action == 'addgruposistemas':
            try:
                f = HdGrupoSistemaEquipoForm(request.POST)
                if f.is_valid():
                    grupo = HdGrupoSistemaEquipo(descripcion=f.cleaned_data['descripcion'],
                                         )
                    grupo.save(request)
                    log(u'Adiciono una nuevo Grupo Sistemas/Equipos: %s' % grupo, request, "add")
                    return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})
        if action=='editgruposistemas':
            try:
                grupo = HdGrupoSistemaEquipo.objects.get(pk=int(request.POST['id']))
                form = HdGrupoSistemaEquipoForm(request.POST, request.FILES)
                if form.is_valid():
                    grupo.descripcion =form.cleaned_data['descripcion']
                    grupo.save(request)
                    log(u'Edito Grupo Sistemas/Equipos: %s' % grupo, request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})
        if action == 'delgruposistemas':
            try:
                grupo = HdGrupoSistemaEquipo.objects.get(pk=int(request.POST['id']))
                # if not grupo.tiene_detalle():
                #     incidente.delete()
                grupo.status=False
                grupo.save()
                log(u'Elimino un grupo sistemas: %s' % grupo, request, "del")
                return JsonResponse({"result": "ok"})
                # else:
                #     return JsonResponse({"result": "ok", "mensaje": u"No se puede eliminar, el registro ese encuentra activo.."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})


        if action == 'addbien':
            try:
                f = HdBienForm(request.POST)
                if f.is_valid():
                    bien = HdBien(
                                   ubicacion_id=int(request.POST['ubicacion']),
                                   gruposistema=HdGrupoSistemaEquipo.objects.get(pk=int(f.cleaned_data['gruposistema'])),
                                   sistemaequipo=f.cleaned_data['sistemaequipo'],
                                   cantidad=f.cleaned_data['cantidad'],
                                   observacion=f.cleaned_data['observacion'],
                                                 )
                    bien.save(request)
                    log(u'Adiciono un Bien: %s' % bien, request, "add")
                    return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})
        if action == 'editbien':
            try:
                bien = HdBien.objects.get(pk=int(request.POST['id']))
                f = HdBienForm(request.POST)
                if f.is_valid():

                    bien.ubicacion_id = int(request.POST['ubicacion'])
                    bien.gruposistema = HdGrupoSistemaEquipo.objects.get(pk=int(f.cleaned_data['gruposistema']))
                    bien.sistemaequipo = f.cleaned_data['sistemaequipo']
                    bien.cantidad = f.cleaned_data['cantidad']
                    bien.observacion = f.cleaned_data['observacion']
                    bien.save(request)
                    log(u'Edito Grupo Bien: %s' % bien, request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})
        if action == 'delbien':
            try:
                bien = HdBien.objects.get(pk=int(request.POST['id']))
                # if not grupo.tiene_detalle():
                #     incidente.delete()
                bien.status = False
                bien.save()
                log(u'Elimino un Bien: %s' % bien, request, "del")
                return JsonResponse({"result": "ok"})
                # else:
                #     return JsonResponse({"result": "ok", "mensaje": u"No se puede eliminar, el registro ese encuentra activo.."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        if action == 'addconf':
            try:
                f = HdConfFrecuenciaForm(request.POST)
                if f.is_valid():
                    conf = HdConfFrecuencia(

                        duracion=f.cleaned_data['duracion'],
                        cantidad=f.cleaned_data['cantidad'],

                    )
                    conf.save(request)
                    log(u'Adiciono un Configuración Frecuencia: %s' % conf, request, "add")
                    return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})
        if action == 'editconf':
            try:
                conf = HdConfFrecuencia.objects.get(pk=int(request.POST['id']))
                f = HdConfFrecuenciaForm(request.POST)
                if f.is_valid():

                    conf.cantidad = f.cleaned_data['cantidad']
                    conf.duracion = f.cleaned_data['duracion']
                    conf.save(request)
                    log(u'Edito Configuración Frecuencia: %s' % conf, request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})
        if action == 'delconf':
            try:
                conf = HdConfFrecuencia.objects.get(pk=int(request.POST['id']))
                # if not grupo.tiene_detalle():
                #     incidente.delete()
                conf.status = False
                conf.save()
                log(u'Elimino Configuración Frecuencia: %s' % conf, request, "del")
                return JsonResponse({"result": "ok"})
                # else:
                #     return JsonResponse({"result": "ok", "mensaje": u"No se puede eliminar, el registro ese encuentra activo.."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        if action == 'addfrecuencia':
            try:
                f = HdFrecuenciaForm(request.POST)

                if f.is_valid():
                    frecuencia = HdFrecuencia(
                        descripcion=f.cleaned_data['descripcion'],
                        consideracion=f.cleaned_data['consideracion'],
                        frecuencia_id = request.POST['frecuencia'],
                        proceso=f.cleaned_data['proceso'],
                        tipomantenimiento=f.cleaned_data['tipomantenimiento'],
                        gruposistema=HdGrupoSistemaEquipo.objects.get(pk=int(f.cleaned_data['gruposistema'])),
                        bien=HdBien.objects.get(pk=int(f.cleaned_data['bien'])),
                        # frecuencia=HdConfFrecuencia.objects.get(pk=f.cleaned_data['frecuencia']),
                    )
                    frecuencia.save(request)
                    log(u'Adiciono Frecuencia: %s' % frecuencia, request, "add")
                    return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})
        if action == 'editfrecuencia':
            try:
                frecuencia = HdFrecuencia.objects.get(pk=int(request.POST['id']))
                f = HdFrecuenciaForm(request.POST)
                if f.is_valid():
                    frecuencia.descripcion = f.cleaned_data['descripcion']
                    frecuencia.consideracion = f.cleaned_data['consideracion']
                    frecuencia.proceso = f.cleaned_data['proceso']
                    frecuencia.tipomantenimiento = f.cleaned_data['tipomantenimiento']
                    frecuencia.gruposistema_id = int(f.cleaned_data['gruposistema'])
                    frecuencia.bien_id = int(f.cleaned_data['bien'])
                    frecuencia.frecuencia_id = int(request.POST['frecuencia'])
                    frecuencia.save(request)
                    log(u'Editó  Frecuencia: %s' % frecuencia, request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})
        if action == 'delfrecuencia':
            try:
                frecuencia = HdFrecuencia.objects.get(pk=int(request.POST['id']))
                # if not grupo.tiene_detalle():
                #     incidente.delete()
                frecuencia.status = False
                frecuencia.save()
                log(u'Elimino  Frecuencia: %s' % frecuencia, request, "del")
                return JsonResponse({"result": "ok"})
                # else:
                #     return JsonResponse({"result": "ok", "mensaje": u"No se puede eliminar, el registro ese encuentra activo.."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        if action == 'addmantenimiento':
            try:
                f = HdCronogramaMantenimientoForm(request.POST)
                if f.is_valid():
                    cronograma = HdCronogramaMantenimiento(
                        tipomantenimiento=f.cleaned_data['tipomantenimiento'],
                        gruposistema=HdGrupoSistemaEquipo.objects.get(pk=int(f.cleaned_data['gruposistema'])),
                        proveedor_id=int(f.cleaned_data['proveedor']) if f.cleaned_data['proveedor'] else None,
                        desde=f.cleaned_data['desde'] if f.cleaned_data['desde'] else None,
                        hasta=f.cleaned_data['hasta'] if f.cleaned_data['hasta'] else None
                    )
                    cronograma.save(request)
                    log(u'Adiciono Cronograma Mantenimiento: %s' % cronograma, request, "add")
                    return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})
        if action == 'editmantenimiento':
            try:
                cronograma = HdCronogramaMantenimiento.objects.get(pk=int(request.POST['id']))
                f = HdCronogramaMantenimientoForm(request.POST)
                if f.is_valid():
                    cronograma.tipomantenimiento = f.cleaned_data['tipomantenimiento']
                    cronograma.gruposistema_id = int(f.cleaned_data['gruposistema'])
                    if f.cleaned_data['proveedor']:
                        cronograma.proveedor_id = int(f.cleaned_data['proveedor'])
                    cronograma.desde = f.cleaned_data['desde']
                    cronograma.hasta = f.cleaned_data['hasta']
                    cronograma.save(request)
                    log(u'Editó  Cronograma Mantenimiento: %s' % cronograma, request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})
        if action == 'delmantenimiento':
            try:
                cronograma = HdCronogramaMantenimiento.objects.get(pk=int(request.POST['id']))
                # if not grupo.tiene_detalle():
                #     incidente.delete()
                cronograma.status = False
                cronograma.save()
                log(u'Elimino  Cronograma Mantenimiento: %s' % cronograma, request, "del")
                return JsonResponse({"result": "ok"})
                # else:
                #     return JsonResponse({"result": "ok", "mensaje": u"No se puede eliminar, el registro ese encuentra activo.."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        if action == 'addreparacion':
            try:
                f = HdReparacionForm(request.POST)
                if f.is_valid():
                    reparacion = HdReparacion(
                        ubicacion_id=int(request.POST['ubicacion']),
                        gruposistema=HdGrupoSistemaEquipo.objects.get(pk=int(f.cleaned_data['gruposistema'])),

                    )
                    reparacion.save(request)
                    log(u'Adiciono Reparación: %s' % reparacion, request, "add")
                    return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})
        if action == 'editreparacion':
            try:
                reparacion = HdReparacion.objects.get(pk=int(request.POST['id']))
                f = HdReparacionForm(request.POST)
                if f.is_valid():

                    reparacion. ubicacion_id = int(request.POST['ubicacion'])
                    reparacion.gruposistema_id = int(f.cleaned_data['gruposistema'])
                    reparacion.save(request)
                    log(u'Editó  Reparación: %s' % reparacion, request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})
        if action == 'delreparacion':
            try:
                reparacion = HdReparacion.objects.get(pk=int(request.POST['id']))
                # if not grupo.tiene_detalle():
                #     incidente.delete()
                reparacion.status = False
                reparacion.save()
                log(u'Elimino  Reparación: %s' % reparacion, request, "del")
                return JsonResponse({"result": "ok"})
                # else:
                #     return JsonResponse({"result": "ok", "mensaje": u"No se puede eliminar, el registro ese encuentra activo.."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        if action == 'adddetalle':
            try:
                f = HdDetCronogramaMantenimientoForm(request.POST)
                if f.is_valid():
                    detmantenimiento = HdDetCronogramaMantenimiento(
                        cronograma=HdCronogramaMantenimiento.objects.get(pk=int(request.POST['idcronograma'])),
                        ubicacion_id=int(request.POST['ubicacion']),
                        bien=HdBien.objects.get(pk=int(f.cleaned_data['bien'])),
                        inventario=f.cleaned_data['inventario'],
                        mes=f.cleaned_data['mes'],
                        descripcion=f.cleaned_data['descripcion'],
                        cantidad=f.cleaned_data['cantidad'],
                    )
                    detmantenimiento.save(request)
                    log(u'Adiciono Detalle Mantenimiento: %s' % detmantenimiento, request, "add")
                    return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})
        if action == 'editdetalle':
            try:
                cronograma = HdDetCronogramaMantenimiento.objects.get(pk=int(request.POST['id']))
                f = HdDetCronogramaMantenimientoForm(request.POST)
                if f.is_valid():
                    cronograma.ubicacion_id = int(request.POST['ubicacion'])
                    cronograma.bien = HdBien.objects.get(pk=int(f.cleaned_data['bien']))
                    cronograma.inventario = f.cleaned_data['inventario']
                    cronograma.mes = f.cleaned_data['mes']
                    cronograma.descripcion = f.cleaned_data['descripcion']
                    cronograma.cantidad = f.cleaned_data['cantidad']
                    cronograma.save(request)
                    log(u'Editó  Detalle Mantenimiento : %s' % cronograma, request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})
        if action == 'deldetalle':
            try:
                detalle = HdDetCronogramaMantenimiento.objects.get(pk=int(request.POST['id']))
                # if not grupo.tiene_detalle():
                #     incidente.delete()
                detalle.status = False
                detalle.save()
                log(u'Elimino  Detalle Mantenimiento: %s' % detalle, request, "del")
                return JsonResponse({"result": "ok"})
                # else:
                #     return JsonResponse({"result": "ok", "mensaje": u"No se puede eliminar, el registro ese encuentra activo.."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        if action == 'adddetreparacion':
            try:
                f = HdDetalle_ReparacionForm(request.POST)
                if f.is_valid():
                    detreparacion = HdDetalle_Reparacion(
                        reparacion=HdReparacion.objects.get(pk=int(request.POST['idreparacion'])),
                        bien=HdBien.objects.get(pk=int(f.cleaned_data['bien'])),
                        descripcion=f.cleaned_data['descripcion'],
                        cantidad=f.cleaned_data['cantidad'],
                    )
                    detreparacion.save(request)
                    log(u'Adiciono Detalle Reparación: %s' % detreparacion, request, "add")
                    return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})
        if action == 'editdetreparacion':
            try:
                detreparacion = HdDetalle_Reparacion.objects.get(pk=int(request.POST['id']))
                f = HdDetalle_ReparacionForm(request.POST)
                if f.is_valid():
                    detreparacion.bien = HdBien.objects.get(pk=int(f.cleaned_data['bien']))
                    detreparacion.descripcion = f.cleaned_data['descripcion']
                    detreparacion.cantidad = f.cleaned_data['cantidad']
                    detreparacion.save(request)
                    log(u'Editó  Detalle Reparación : %s' % detreparacion, request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})
        if action == 'deldetreparacion':
            try:
                detreparacion = HdDetalle_Reparacion.objects.get(pk=int(request.POST['id']))
                # if not grupo.tiene_detalle():
                #     incidente.delete()
                detreparacion.status = False
                detreparacion.save()
                log(u'Elimino  Detalle Reparación: %s' % detreparacion, request, "del")
                return JsonResponse({"result": "ok"})
                # else:
                #     return JsonResponse({"result": "ok", "mensaje": u"No se puede eliminar, el registro ese encuentra activo.."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        if action == 'addplan':
            try:
                f = HdFrecuenciaForm(request.POST)
                if f.is_valid():
                    frecuencia = HdFrecuencia(

                        descripcion=f.cleaned_data['descripcion'],
                        consideracion=f.cleaned_data['consideracion'],
                        proceso=f.cleaned_data['proceso'],
                        tipomantenimiento=f.cleaned_data['tipomantenimiento'],
                        gruposistema=int(f.cleaned_data['gruposistema']),
                        bien_id=int(f.cleaned_data['bien']),
                        frecuencia_id=int(f.cleaned_data['frecuencia']),
                    )
                    frecuencia.save(request)
                    log(u'Adiciono Frecuencia: %s' % frecuencia, request, "add")
                    return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})
        if action == 'editplan':
            try:
                frecuencia = HdConfFrecuencia.objects.get(pk=int(request.POST['id']))
                f = HdConfFrecuenciaForm(request.POST)
                if f.is_valid():
                    frecuencia.descripcion = f.cleaned_data['descripcion']
                    frecuencia.consideracion = f.cleaned_data['consideracion']
                    frecuencia.proceso = f.cleaned_data['proceso']
                    frecuencia.tipomantenimiento = f.cleaned_data['tipomantenimiento']
                    frecuencia.gruposistema = int(f.cleaned_data['gruposistema'])
                    frecuencia.bien_id = int(f.cleaned_data['bien'])
                    frecuencia.frecuencia_id = int(f.cleaned_data['frecuencia'])
                    frecuencia.save(request)
                    log(u'Edito  Frecuencia: %s' % frecuencia, request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})
        if action == 'delplan':
            try:
                frecuencia = HdFrecuencia.objects.get(pk=int(request.POST['id']))
                # if not grupo.tiene_detalle():
                #     incidente.delete()
                frecuencia.status = False
                frecuencia.save()
                log(u'Elimino  Frecuencia: %s' % frecuencia, request, "del")
                return JsonResponse({"result": "ok"})
                # else:
                #     return JsonResponse({"result": "ok", "mensaje": u"No se puede eliminar, el registro ese encuentra activo.."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        if action == 'addmantmaterial':
            try:
                f = HdMaterialMantenimientoForm(request.POST)
                if f.is_valid():
                    mant=mantmaterial = HdMaterialMantenimiento(
                        ubicacion_id=int(request.POST['ubicacion']),
                        tipomantenimiento=f.cleaned_data['tipomantenimiento'],
                        gruposistema=HdGrupoSistemaEquipo.objects.get(pk=int(f.cleaned_data['gruposistema'])),
                        bien=HdBien.objects.get(pk=int(f.cleaned_data['bien'])),
                        tipobien=f.cleaned_data['tipobien'],
                        proceso=f.cleaned_data['proceso'],
                    )
                    mantmaterial.save(request)
                    listaayudantes = HdDetalle_Grupo.objects.filter(pk__in=f.cleaned_data['ayudantes'])
                    if listaayudantes:
                        for ayudante in listaayudantes:
                            detalleayudantes = HdMaterialMantenimiento_Responsable(materialmantenimiento=mant,
                                                                             agente=ayudante
                                                                             )
                            detalleayudantes.save(request)
                    log(u'Adiciono  Mantenimiento Material: %s' % mantmaterial, request, "add")
                    return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})
        if action == 'editmantmaterial':
            try:
                mantmaterial = HdMaterialMantenimiento.objects.get(pk=int(request.POST['id']))
                f = HdMaterialMantenimientoForm(request.POST)
                if f.is_valid():
                    # mantmaterial.ubicacion_id=int(request.POST['ubicacion'])
                    # mantmaterial.tipomantenimiento=f.cleaned_data['tipomantenimiento']
                    # mantmaterial.gruposistema=HdGrupoSistemaEquipo.objects.get(pk=int(f.cleaned_data['gruposistema']))
                    # mantmaterial.bien=HdBien.objects.get(pk=int(f.cleaned_data['bien']))
                    # mantmaterial.tipobien=f.cleaned_data['tipobien']
                    # mantmaterial.proceso=f.cleaned_data['proceso']
                    # mantmaterial.save(request)

                    detmat=HdMaterialMantenimiento_Responsable.objects.filter(status=True,materialmantenimiento=(mantmaterial.pk))
                    detmat.delete()
                    listaayudantes = HdDetalle_Grupo.objects.filter(pk__in=f.cleaned_data['ayudantes'])
                    if listaayudantes:
                        for ayudante in listaayudantes:

                            detalleayudantes = HdMaterialMantenimiento_Responsable(materialmantenimiento=mantmaterial,
                                                                                   agente=ayudante
                                                                                   )
                            detalleayudantes.save(request)
                    log(u'Editó   Mantenimiento Material: %s' % mantmaterial, request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})
        if action == 'delmantmaterial':
            try:
                cronograma = HdMaterialMantenimiento.objects.get(pk=int(request.POST['id']))
                # if not grupo.tiene_detalle():
                #     incidente.delete()
                cronograma.status = False
                cronograma.save()
                log(u'Elimino  Cronograma Mantenimiento: %s' % cronograma, request, "del")
                return JsonResponse({"result": "ok"})
                # else:
                #     return JsonResponse({"result": "ok", "mensaje": u"No se puede eliminar, el registro ese encuentra activo.."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        if action == 'selectubicacion':
            try:
                if 'id' in request.POST:
                    lista = []
                    ubicaciones = HdBloqueUbicacion.objects.filter(bloque_id=int(request.POST['id']), status=True)
                    for ubi in ubicaciones:
                        lista.append([ubi.id, ubi.ubicacion.nombre])
                    return JsonResponse({"result": "ok", 'lista': lista})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al consulatar los datos."})
        if action == 'selectbloquerep':
            try:

                    lista = []
                    ubicaciones = HdReparacion.objects.filter( status=True,usuario_creacion=int(persona.usuario.id))
                    for ubi in ubicaciones:
                        lista.append([ubi.ubicacion.bloque.pk, ubi.ubicacion.bloque.nombre])
                    return JsonResponse({"result": "ok", 'lista': lista})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al consulatar los datos."})
        if action == 'selectubicacionrep':
            try:
                if 'id' in request.POST:
                    lista = []
                    ubicaciones = HdReparacion.objects.filter(status=True,ubicacion__bloque__pk=int(request.POST['id']),usuario_creacion=int(persona.usuario.id))
                    for ubi in ubicaciones:
                        lista.append([ubi.ubicacion.ubicacion.pk, ubi.ubicacion.ubicacion.nombre])
                    return JsonResponse({"result": "ok", 'lista': lista})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al consulatar los datos."})
        if action == 'selectgruporeparacion':
            try:
                if 'id' in request.POST and 'ubi' in request.POST:
                    lista = []
                    ubicaciones = HdReparacion.objects.filter(status=True,ubicacion__bloque__pk=int(request.POST['id']),ubicacion__ubicacion__pk=int(request.POST['ubi']),usuario_creacion=int(persona.usuario.id))
                    for ubi in ubicaciones:
                        lista.append([ubi.gruposistema.pk, ubi.gruposistema.descripcion])
                    return JsonResponse({"result": "ok", 'lista': lista})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al consulatar los datos."})
        if action == 'selecttecnico':
            try:

                lista = []
                ubicaciones = HdDetalle_Reparacion.objects.values_list('usuario_creacion').filter(status=True)
                ubica = Persona.objects.filter(status=True,usuario__pk__in=ubicaciones)
                for ubi in ubica:
                    lista.append([ubi.pk, ubi.nombre_completo_inverso()])
                return JsonResponse({"result": "ok", 'lista': lista})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al consulatar los datos."})
        if action == 'selectbienes':
            try:
                if 'id' in request.POST:
                    lista = []
                    bien = HdMaterialMantenimiento.objects.filter(gruposistema_id=int(request.POST['id']), status=True).distinct('tipobien')

                    for ubi in bien:
                        lista.append([ubi.bien.pk, ubi.bien.sistemaequipo])
                    return JsonResponse({"result": "ok", 'listas': lista})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al consulatar los datos."})
        if action == 'selectvalor':
            try:
                requerido =0.0
                if 'id' in request.POST and 'grupo' in request.POST :
                    lista = []
                    pre = HdMaterialMantenimiento_Material.objects.filter(materialmantenimiento__gruposistema=int(request.POST['grupo']),materialmantenimiento__bien=int(request.POST['id']), status=True)

                    for ubi in pre:
                         requerido=requerido+float(ubi.total)
                    iva=float(requerido)*0.12
                    total=requerido+iva
                    lista.append([requerido, iva,total])
                    return JsonResponse({"result": "ok", 'listas': lista})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al consulatar los datos."})

        if action == 'selectcronograma':
            try:
                if 'id' in request.POST and 'blo' in request.POST:
                    lista = []

                    cronogramasem=HdCronogramaMantenimientoSem.objects.get(status=True,pk=int(request.POST['id']))
                    ubicaciones = HdDetCronogramaMantenimiento.objects.filter(cronograma__gruposistema=int(cronogramasem.gruposistema.pk),mes=int(cronogramasem.mes),ubicacion__bloque__pk=int(request.POST['blo']), status=True).distinct('ubicacion__ubicacion')

                    for ubi in ubicaciones:
                        lista.append([ubi.ubicacion.id, ubi.ubicacion.ubicacion.nombre])


                    return JsonResponse({"result": "ok", 'lista': lista})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al consulatar los datos."})
        if action == 'selectbien':
            try:
                if 'sem' in request.POST  and 'blo' in request.POST :
                    listas = []
                    cronogramasem = HdCronogramaMantenimientoSem.objects.get(status=True, pk=int(request.POST['sem']))
                    bien = HdDetCronogramaMantenimiento.objects.filter(mes=int(cronogramasem.mes),status=True,ubicacion__pk=int(request.POST['blo']))
                    for ubi in bien:
                        listas.append([ubi.bien.pk, ubi.bien.sistemaequipo, ubi.cantidad])
                    return JsonResponse({"result": "ok", 'listas': listas})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al consulatar los datos."})

        if action == 'selectbiencantidad':
            try:
                listas = []
                if 'sem' in request.POST and 'blo' in request.POST and  'id' in request.POST:

                    cronogramasem = HdCronogramaMantenimientoSem.objects.get(status=True, pk=int(request.POST['sem']))
                    bien = HdDetCronogramaMantenimiento.objects.filter(mes=int(cronogramasem.mes), status=True,
                                                                       ubicacion__pk=int(request.POST['blo']),bien__pk=int(request.POST['id']))
                    for ubi in bien:
                        listas.append([ ubi.cantidad])

                    return JsonResponse({"result": "ok", 'listas': listas})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al consulatar los datos."})
        if action == 'selectcronogramabloque':
            try:
                if 'sem' in request.POST:
                    lista = []

                    cronogramasem=HdCronogramaMantenimientoSem.objects.get(status=True,pk=int(request.POST['sem']))
                    ubicaciones = HdDetCronogramaMantenimiento.objects.filter(cronograma__gruposistema=int(cronogramasem.gruposistema.pk),mes=int(cronogramasem.mes), status=True).distinct('ubicacion__bloque')
                    for ubi in ubicaciones:
                        lista.append([ubi.ubicacion.bloque.pk, ubi.ubicacion.bloque.nombre])

                    return JsonResponse({"result": "ok", 'lista': lista})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al consulatar los datos."})

        if action == 'selectgrupopresupuesto':
            try:
                lista = []
                if HdMaterialMantenimiento.objects.filter(status=True).exists():

                    ubicaciones = HdMaterialMantenimiento.objects.filter(status=True).distinct('gruposistema')
                    for ubi in ubicaciones:
                        lista.append([ubi.gruposistema.pk, ubi.gruposistema.descripcion])

                return JsonResponse({"result": "ok", 'lista': lista})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al consulatar los datos."})
        if action == 'addmateriales':
            try:
                mantmaterial = HdMaterialMantenimiento_Material(
                                                            materialmantenimiento_id=request.POST['idmaterialmantenimiento'],
                                                            material=request.POST['idmaterial'],
                                                            unidadmedida_id=request.POST['idunidadmedida'],
                                                            precio=request.POST['idprecio'],
                                                            cantidad=request.POST['idcantidad'],
                                                            total=request.POST['idtotal'])
                mantmaterial.save(request)
                log(u'Adiciono nuevo Material Mantenimiento: %s' % mantmaterial, request, "add")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": "Error al guardar los datos."})

        if action == 'vermantmaterial':
            try:
                lista = []
                mantmaterial = HdMaterialMantenimiento.objects.get(pk=int(request.POST['idmaterialmantenimiento']))
                data['mantmateriales'] = HdMaterialMantenimiento_Material.objects.filter(status=True,materialmantenimiento=int(mantmaterial.pk))
                template = get_template("helpdesk_hdplanificacion/consultamantmaterial.html")
                json_content = template.render(data)
                return JsonResponse(
                    {"result": "ok", 'html': json_content, 'title': u'Seleccionar el periodo academico'})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al generar excel."})

        if action == 'conmateriales':
            try:
                material = HdMaterialMantenimiento_Material.objects.get(pk=request.POST['idmater'], status=True)
                nombre = material.material
                return JsonResponse({"result": "ok", "nombre": nombre})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": "Error al eliminar los datos."})

        if action == 'delmateriales':
            try:
                if HdMaterialMantenimiento_Material.objects.filter(pk=request.POST['idmaterial'], status=True):
                    materialincidente = HdMaterialMantenimiento_Material.objects.get(pk=request.POST['idmaterial'], status=True)
                    materialincidente.delete()
                    log(u'Elimino material : %s' % materialincidente, request, "del")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": "Error al eliminar los datos."})

        if action == 'addcronograma':
            try:
                f = HdCronogramaMantenimientoSemForm(request.POST)
                if f.is_valid():
                    cronograma = HdCronogramaMantenimientoSem(
                        mes=f.cleaned_data['mes'],
                        gruposistema=HdGrupoSistemaEquipo.objects.get(pk=int(f.cleaned_data['gruposistema'])),

                    )
                    cronograma.save(request)
                    log(u'Adiciono Cronograma Mantenimiento Semanal: %s' % cronograma, request, "add")
                    return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})



        if action == 'addactividad':
            try:
                detcronograma = HdDetCronogramaMantenimientoSem(
                    cronograma_id=request.POST['idmaterialmantenimiento'],
                    bloque_id=request.POST['idubicacion'],
                    bien_id=request.POST['idbien'],
                    fechainicio=request.POST['idfechainicio'],
                    cantidad=request.POST['idcantidad'],
                    fechafin=request.POST['idfechafin'])

                detcronograma.save(request)
                log(u'Adiciono nuevo Actividad Semanal: %s' % detcronograma, request, "add")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": "Error al guardar los datos."})

        if action == 'veractividad':
            try:
                lista = []
                mantcronograma = HdCronogramaMantenimientoSem.objects.get(pk=int(request.POST['idmaterialmantenimiento']))
                data['mantcronograma'] = HdDetCronogramaMantenimientoSem.objects.filter(status=True,
                                                                                         cronograma=int(
                                                                                             mantcronograma.pk))
                template = get_template("helpdesk_hdplanificacion/consultaactividad.html")
                json_content = template.render(data)
                return JsonResponse(
                    {"result": "ok", 'html': json_content, 'title': u'Seleccionar el periodo academico'})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al generar excel."})

        if action == 'conactividad':
            try:
                actividad = HdDetCronogramaMantenimientoSem.objects.get(pk=request.POST['idmater'], status=True)
                nombre = actividad.bien.sistemaequipo
                return JsonResponse({"result": "ok", "nombre": nombre})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": "Error al eliminar los datos."})

        if action == 'delactividad':
            try:
                if HdDetCronogramaMantenimientoSem.objects.filter(pk=request.POST['idactividad'], status=True):
                    actividad = HdDetCronogramaMantenimientoSem.objects.get(pk=request.POST['idactividad'],status=True)
                    actividad.status=False
                    actividad.save()
                    log(u'Elimino Actividades Semanales : %s' % actividad, request, "del")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": "Error al eliminar los datos."})
        if action == 'editcronograma':
            try:
                crono = HdCronogramaMantenimientoSem.objects.get(pk=int(request.POST['id']))
                f = HdCronogramaMantenimientoSemForm(request.POST)
                # if f.is_valid():
                #     crono.descripcion = f.cleaned_data['descripcion']
                #     crono.gruposistema = int(f.cleaned_data['gruposistema'])

                    # crono.save(request)
                log(u'Adicionó Actividades: %s' % crono, request, "edit")
                return JsonResponse({"result": "ok"})

            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        if action == 'addpresupuesto':
            try:
                f = HdPresupuestoRecursoForm(request.POST)

                if f.is_valid():
                    presupuesto = HdPresupuestoRecurso(
                        presupuestoiva=f.cleaned_data['presupuestoiva'],
                        presupuestoreq=f.cleaned_data['presupuestoreq'],
                        presupuestototal=f.cleaned_data['presupuestototal'],
                        # tipobien=f.cleaned_data['tipobien'],
                        gruposistema_id=int(request.POST['gruposistema']),
                        bien_id=int(request.POST['bien']),
                        # frecuencia=HdConfFrecuencia.objects.get(pk=f.cleaned_data['frecuencia']),
                    )
                    presupuesto.save(request)
                    log(u'Adicionó Presupuesto Recurso: %s' % presupuesto, request, "add")
                    return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})
        if action == 'editpresupuesto':
            try:
                presupuesto = HdPresupuestoRecurso.objects.get(pk=int(request.POST['id']))
                f = HdPresupuestoRecursoForm(request.POST)
                if f.is_valid():
                    presupuesto.presupuestoiva = f.cleaned_data['presupuestoiva']
                    # presupuesto.tipobien = f.cleaned_data['tipobien']
                    presupuesto.presupuestoreq = f.cleaned_data['presupuestoreq']
                    presupuesto.presupuestototal = f.cleaned_data['presupuestototal']
                    presupuesto.gruposistema_id = int(request.POST['gruposistema'])
                    presupuesto.bien_id = int(request.POST['bien'])
                    presupuesto.save(request)
                    log(u'Editó   Presupuesto Recurso: %s' % presupuesto, request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})
        if action == 'delpresupuesto':
            try:
                presupuesto = HdPresupuestoRecurso.objects.get(pk=int(request.POST['id']))
                # if not grupo.tiene_detalle():
                #     incidente.delete()
                presupuesto.status = False
                presupuesto.save()
                log(u'Elimino  Presupuesto Recurso: %s' % presupuesto, request, "del")
                return JsonResponse({"result": "ok"})
                # else:
                #     return JsonResponse({"result": "ok", "mensaje": u"No se puede eliminar, el registro ese encuentra activo.."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})
        elif action == 'afmarcaymodelo':
            try:
                listatareas = ActivoTecnologico.objects.filter(id=request.POST['id_tipo'], status=True)
                lista = []
                for lis in listatareas:
                    lista.append([lis.activotecnologico.marca, lis.activotecnologico.modelo,lis.activotecnologico.responsable.nombre_completo(),lis.tipoactivo.pk])
                return JsonResponse({'result': 'ok', 'lista': lista})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})
        if action == 'listatipopieza':
            try:
                listatareas = HdPiezaPartes.objects.filter(grupocategoria_id=request.POST['idcat'], status=True)
                lista = []
                for lis in listatareas:
                    lista.append([lis.id, lis.descripcion])
                return JsonResponse({'result': 'ok', 'lista': lista})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})
        elif action == 'listatipomantenimientosgdanio':
            try:
                listatareas = HdMantenimientoGruDanios.objects.filter(grupocategoria_id=request.POST['id_tipo'], status=True, activo=True)
                lista = []
                for lis in listatareas:
                    lista.append([lis.id, lis.descripcion])
                return JsonResponse({'result': 'ok', 'lista': lista})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})
        elif action == 'listatipomantenimiento':
            try:
                listatareas = HdMantenimientoGruCategoria.objects.filter(grupocategoria_id=request.POST['id_tipo'], status=True, activo=True)
                lista = []
                for lis in listatareas:
                    lista.append([lis.id, lis.descripcion])
                return JsonResponse({'result': 'ok', 'lista': lista})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})
        elif action == 'detalle_mantenimientopreven':
            try:
                data['mantenimiento'] = mantenimiento = HdDetMantenimientosActivos.objects.get(pk=encrypt(request.POST['id']))
                data['tareasmantenimiento'] = HdMantenimientoGruCategoria.objects.filter(grupocategoria=mantenimiento.tipoactivoc, status=True,activo=True)
                data['tareasactivo'] = HdTareasActivosPreventivos.objects.values_list('grupos_id', flat=True).filter(mantenimiento=mantenimiento, status=True)
                data['ppmantenimiento'] = HdPiezaPartes.objects.filter(grupocategoria=mantenimiento.tipoactivoc,status=True)
                data['piezaparteactivo'] = HdPiezaParteActivosPreventivos.objects.filter(mantenimiento=mantenimiento,status=True)
                data['daniomantenimiento'] = HdMantenimientoGruDanios.objects.filter(grupocategoria=mantenimiento.tipoactivoc, status=True,activo=True)
                data['danioactivo'] = HdTareasActivosPreventivosDanios.objects.filter(mantenimiento=mantenimiento, status=True)
                data['estdan'] = ESTADO_DANIO
                template = get_template("helpdesk_hdplanificacion/modal/detallemantenimientopreventivo.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'html': json_content})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})
        elif action == 'addmantenimientoactivo':
            try:
                form = HdMantenimientosActivosForm(request.POST, request.FILES)
                fechadefinida = '2022-10-18'
                if 'archivo' in request.FILES:
                    newfile = request.FILES['archivo']
                    if newfile:
                        newfilesd = newfile._name
                        ext = newfilesd[newfilesd.rfind("."):]
                        if not ext.lower() == '.pdf':
                            return JsonResponse({"result": "bad", "mensaje": u"Error, solo archivos .pdf."})
                        if newfile.size > 2194304:
                            return JsonResponse({"result": "bad", "mensaje": u"Error, archivo mayor a 2 Mb."})
                        if newfile:
                            newfile._name = generar_nombre("archivo_malla_", newfile._name)
                if form.is_valid():
                    id_cronograma = request.POST['cronograma']
                    id_cpu = request.POST['activotecno']
                    id_tipact = request.POST['tipoactivo']
                    estusu = form.cleaned_data['estusu']
                    fecha = form.cleaned_data['fecha']
                    horamax = form.cleaned_data['horamax']
                    minutomax = form.cleaned_data['minutomax']
                    estfrec = form.cleaned_data['estfrec']
                    estfent = form.cleaned_data['estfent']
                    marca = form.cleaned_data['marca']
                    modelo = form.cleaned_data['modelo']
                    bsugiere = form.cleaned_data['bsugiere']
                    dsugiere = form.cleaned_data['dsugiere']
                    observacion = form.cleaned_data['observacion']
                    limp = request.POST['id_limp']
                    # danio = request.POST['id_dani']
                    piezaparte = request.POST.getlist('piezaparte[]')
                    caracteristica = request.POST.getlist('catacteristica[]')
                    danio = request.POST.getlist('danio[]')
                    daniodes = request.POST.getlist('danioop[]')

                    verificarmantenimiento = HdDetMantenimientosActivos.objects.filter(activotecno_id=id_cpu, fecha__year=fecha.year, status=True).count()
                    if verificarmantenimiento < 2:
                        mantenimientosactivos = HdDetMantenimientosActivos(activotecno_id=id_cpu,
                                                                           tipoactivoc_id=id_tipact,
                                                                           cronograma_id=id_cronograma,
                                                                           estusu=estusu,
                                                                           fecha=fecha,
                                                                           horamax=horamax,
                                                                           minutomax=minutomax,
                                                                           funcionarecibe=estfrec,
                                                                           funcionaentrega=estfent,
                                                                           marca=marca,
                                                                           modelo=modelo,
                                                                           sbequipo=bsugiere,
                                                                           descbaja=dsugiere,
                                                                           observaciones=observacion,
                                                                           nuevo=True,
                                                                           persona=persona)
                        mantenimientosactivos.save(request)
                        limp = limp.split(',')
                        for elemento in limp:
                            detalle = HdTareasActivosPreventivos(mantenimiento=mantenimientosactivos,
                                                               grupos_id=elemento)
                            detalle.save(request)
                        c = 0
                        while c < len(danio):
                            detalle = HdTareasActivosPreventivosDanios(mantenimiento=mantenimientosactivos,
                                                                     grupos_id=int(HdMantenimientoGruDanios.objects.get(descripcion=danio[c], grupocategoria_id=id_tipact, status=True).id),
                                                                     estadodanio=daniodes[c])
                            detalle.save(request)
                            c += 1
                        counter = 0
                        while counter < len(piezaparte):
                            detalle = HdPiezaParteActivosPreventivos(mantenimiento=mantenimientosactivos,
                                                                   piezaparte_id=int(HdPiezaPartes.objects.get(descripcion=piezaparte[counter], grupocategoria_id=id_tipact, status=True).id),
                                                                   descripcion=caracteristica[counter])
                            detalle.save(request)
                            counter += 1
                        if 'archivo' in request.FILES:
                            newfile = request.FILES['archivo']
                            ext = newfile.name.split('.')[-1]
                            if not ext in ['pdf','jpg','jpeg','png','word']:
                                raise NameError(u'Incorrecto formato de archivo')
                            if newfile:
                                newfile._name = generar_nombre("evidencia_comunicado", newfile._name)
                                mantenimientosactivos.archivo = newfile
                                mantenimientosactivos.save(request)
                        confmantenimiento = SolicitudConfirmacionMantenimiento(
                            mantenimiento = mantenimientosactivos,
                            observacion = mantenimientosactivos.observaciones,
                            estado = 1
                        )
                        confmantenimiento.save(request)
                        log(u'Agrego registro de confirmacion de mantenimiento %s'%(confmantenimiento.__str__()),request,"add")
                        historial =  HistorialSolicitudConfirmacionMantenimiento(
                            solicitud = confmantenimiento,
                            observacion = confmantenimiento.observacion,
                            estado = confmantenimiento.estado,
                            persona = persona
                        )
                        historial.save(request)
                        log(u'Agrego registro de historial de confirmacion de mantenimiento %s' % (historial.__str__()), request, "add")
                        if str(mantenimientosactivos.fecha) >= fechadefinida:
                            titulo = "Confirmación de mantenimiento"
                            cuerpo = "Mantenimiento preventivo realizado por: %s" % (mantenimientosactivos.persona)
                            notificacion(titulo,
                                         cuerpo, mantenimientosactivos.activotecno.activotecnologico.responsable, None, 'th_hojavida', mantenimientosactivos.pk,
                                         1, 'sga', mantenimientosactivos, request)
                            notificacion(titulo,
                                         cuerpo, mantenimientosactivos.activotecno.activotecnologico.responsable, None, 'th_hojavida', mantenimientosactivos.pk,
                                         1, 'sagest', mantenimientosactivos, request)
                        else:
                            solicitud = SolicitudConfirmacionMantenimiento.objects.get(pk=confmantenimiento.id)
                            solicitud.estado = 2
                            solicitud.save(request)
                            log(u'Confirmacion de mantenimiento automatica: %s' % solicitud, request, "add")
                            historial = HistorialSolicitudConfirmacionMantenimiento(
                                solicitud=solicitud,
                                observacion=solicitud.observacion,
                                estado=solicitud.estado,
                                persona=persona
                            )
                            historial.save(request)
                            log(u'Agrego registro de historial de confirmacion automatica %s' % (historial.__str__()),
                                request, "add")
                        log(u'Asigno mantenimiento preventivo Activo: %s' % mantenimientosactivos, request, "add")
                        return JsonResponse({"result": "ok"})
                    else:
                        return JsonResponse({"result": "bad", "mensaje": "El activo cuenta con 2 mantenimientos correspondiente al año de la fecha seleccionada"})
                else:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': "bad", "form": [{k: v[0]} for k, v in form.errors.items()],
                                         "message": "Error en el formulario"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'editmantenimientoactivo':
            try:
                form = HdMantenimientosActivosForm(request.POST, request.FILES)
                if form.is_valid():
                    eli_arc = request.POST['archivo-clear'] if 'archivo-clear' in request.POST else ''
                    id_activoman = request.POST['id']
                    id_cpu = request.POST['activotecno']
                    id_tipact = request.POST['tipoactivo']
                    estusu = form.cleaned_data['estusu']
                    fecha = form.cleaned_data['fecha']
                    horamax = form.cleaned_data['horamax']
                    minutomax = form.cleaned_data['minutomax']
                    estfrec = form.cleaned_data['estfrec']
                    estfent = form.cleaned_data['estfent']
                    marca = form.cleaned_data['marca']
                    modelo = form.cleaned_data['modelo']
                    bsugiere = form.cleaned_data['bsugiere']
                    dsugiere = form.cleaned_data['dsugiere']
                    observacion = form.cleaned_data['observacion']
                    limp = request.POST['id_limpiar']
                    # danio = request.POST['id_danio']
                    piezaparte = request.POST.getlist('piezaparte[]')
                    caracteristica = request.POST.getlist('catacteristica[]')
                    danio = request.POST.getlist('danio[]')
                    daniodes = request.POST.getlist('danioop[]')

                    mantenimientosactivos = HdDetMantenimientosActivos.objects.get(id=id_activoman)
                    mantenimientosactivos.activofijo_id = id_cpu
                    mantenimientosactivos.tipoactivoc_id = id_tipact
                    mantenimientosactivos.estusu = estusu
                    mantenimientosactivos.fecha = fecha
                    mantenimientosactivos.horamax = horamax
                    if eli_arc == 'on':
                        mantenimientosactivos.archivo = None
                    mantenimientosactivos.minutomax = minutomax
                    mantenimientosactivos.funcionarecibe = estfrec
                    mantenimientosactivos.funcionaentrega = estfent
                    mantenimientosactivos.marca = marca
                    mantenimientosactivos.modelo = modelo
                    mantenimientosactivos.sbequipo = bsugiere
                    mantenimientosactivos.descbaja = dsugiere
                    mantenimientosactivos.observaciones = observacion
                    mantenimientosactivos.save(request)
                    if 'archivo' in request.FILES:
                        newfile = request.FILES['archivo']
                        ext = newfile._name.split('.')[-1]
                        if not ext in ['pdf','jpg','jpeg','png','word']:
                            raise NameError(u'Incorrecto formato de archivo')
                        newfile._name = generar_nombre("evidencia_comunicado", newfile._name)
                        mantenimientosactivos.archivo = newfile
                        mantenimientosactivos.save(request)
                    elementos = limp.split(',')
                    limpiezaquitar = []
                    if len(elementos)>0 and [''] != elementos:
                        limpiezaquitar = HdTareasActivosPreventivos.objects.filter(mantenimiento=mantenimientosactivos).exclude(grupos_id__in=elementos)
                    if [''] == elementos:
                        limpiezaquitar = HdTareasActivosPreventivos.objects.filter(mantenimiento=mantenimientosactivos)
                    for lim in limpiezaquitar:
                        lim.status = False
                        lim.save(request)
                    for elemento in elementos:
                        if elemento != '':
                            if not HdTareasActivosPreventivos.objects.values('id').filter(mantenimiento=mantenimientosactivos,grupos_id=elemento).exists():
                                detalle = HdTareasActivosPreventivos(mantenimiento=mantenimientosactivos,
                                                                   grupos_id=elemento)
                                detalle.save(request)
                            elif HdTareasActivosPreventivos.objects.values('id').filter(mantenimiento=mantenimientosactivos,grupos_id=elemento,status=False).exists():
                                detalle = HdTareasActivosPreventivos.objects.filter(mantenimiento=mantenimientosactivos,grupos_id=elemento,status=False).order_by('-id').first()
                                detalle.status = True
                                detalle.save(request)

                    c = 0
                    daniosexclu = HdTareasActivosPreventivosDanios.objects.filter(status=True,mantenimiento=mantenimientosactivos).exclude(grupos_id__in =danio)
                    for dani in daniosexclu:
                        dani.status = False
                        dani.save(request)
                    while c < len(danio):
                        idgrupomandanio = HdMantenimientoGruDanios.objects.get(id=danio[c], grupocategoria_id=id_tipact,status=True).id
                        if not HdTareasActivosPreventivosDanios.objects.values('id').filter(status=True,mantenimiento=mantenimientosactivos,grupos_id =idgrupomandanio).exists():
                            detalle = HdTareasActivosPreventivosDanios(mantenimiento=mantenimientosactivos,
                                                                 grupos_id=int(idgrupomandanio),
                                                                 estadodanio=daniodes[c])
                            detalle.save(request)
                        c += 1
                    piezaparteexclu = HdPiezaParteActivosPreventivos.objects.filter(status=True, mantenimiento=mantenimientosactivos).exclude(piezaparte_id__in=piezaparte)
                    for pieza in piezaparteexclu:
                        pieza.status = False
                        pieza.save(request)
                    counter = 0
                    while counter < len(piezaparte):
                        idpiezaparte = HdPiezaPartes.objects.get(id=piezaparte[counter], status=True).id
                        if not HdPiezaParteActivosPreventivos.objects.values('id').filter(status=True, mantenimiento=mantenimientosactivos, piezaparte_id=idpiezaparte).exists():
                            detalle = HdPiezaParteActivosPreventivos(mantenimiento=mantenimientosactivos,
                                                                   piezaparte_id=int(idpiezaparte),
                                                                   descripcion=caracteristica[counter])
                            detalle.save(request)
                        counter += 1
                    log(u'Editó mantenimiento preventivo Activo: %s' % mantenimientosactivos, request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': "bad", "form": [{k: v[0]} for k, v in form.errors.items()],
                                         "message": "Error en el formulario"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos. %s"%(ex.__str__())})

        elif action == 'addtipobien':
            try:
                form = TipoBienForm(request.POST)
                if form.is_valid():
                    hd_grupos = HdGruposCategoria(
                        descripcion = form.cleaned_data['descripcion']
                    )
                    hd_grupos.save(request)
                    log(u'Agregó un tipo bien: %s'%(hd_grupos.__str__()),request,'add')
                    return JsonResponse({'result':False,'message':'Registro exitosa'})
                else:
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()],
                                         "message": "Error en el formulario"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': f'Error al guardar los datos. {ex}'})

        elif action == 'edittipobien':
            try:
                form = TipoBienForm(request.POST)
                filtro= HdGruposCategoria.objects.get(id = int(encrypt(request.POST['id'])))
                if form.is_valid():
                    filtro.descripcion = form.cleaned_data['descripcion']
                    filtro.save(request)
                    log(u'Editó un tipo bien: %s'%(filtro.__str__()),request,'edit')
                    return JsonResponse({'result':False,'message':'Edicion exitosa'})
                else:
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()],
                                         "message": "Error en el formulario"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': f'Error al guardar los datos. {ex}'})

        elif action == 'deletetipobien':
            try:
                id = int(encrypt(request.POST['id']))
                hd_grupos = HdGruposCategoria.objects.get(id=id)
                hd_grupos.status = False
                hd_grupos.save(request)
                log(u'Eliminó el registro: %s - %s'%(hd_grupos.__str__(),hd_grupos.pk),request,'delete')
                return JsonResponse({'error':False,'mensaje':u'Registro eliminado'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'error': True, 'mensaje': f'Error al guardar los datos. {ex}'})

        elif action == 'deletedetalle':
            try:
                id = int(encrypt(request.POST['id']))
                hd_cronomante = HdDetMantenimientosActivos.objects.get(id=id)
                hd_cronomante.status = False
                hd_cronomante.save(request)
                log(u'Eliminó el registro: %s - %s'%(hd_cronomante.__str__(),hd_cronomante.pk),request,'delete')
                return JsonResponse({'error':False,'mensaje':u'Registro eliminado'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'error': True, 'mensaje': f'Error al guardar los datos. {ex}'})

        elif action == 'addtarealimp':
            try:
                form = TareasMantenimientoForm(request.POST)
                if form.is_valid():
                    hd_grupos = HdMantenimientoGruCategoria(
                        grupocategoria= form.cleaned_data['tipobien'],
                        descripcion = form.cleaned_data['descripcion']
                    )
                    hd_grupos.save(request)
                    log(u'Agregó un mantenimiento de limpieza: %s'%(hd_grupos.__str__()),request,'add')
                    return JsonResponse({'result':False,'message':'Registro exitosa'})
                else:
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()],
                                         "message": "Error en el formulario"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': f'Error al guardar los datos. {ex}'})

        elif action == 'edittarealimp':
            try:
                form = TareasMantenimientoForm(request.POST)
                filtro= HdMantenimientoGruCategoria.objects.get(id = int(encrypt(request.POST['id'])))
                if form.is_valid():
                    filtro.grupocategoria = form.cleaned_data['tipobien']
                    filtro.descripcion = form.cleaned_data['descripcion']
                    filtro.save(request)
                    log(u'Editó un mantenimiento de limpieza: %s'%(filtro.__str__()),request,'edit')
                    return JsonResponse({'result':False,'message':'Edicion exitosa'})
                else:
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()],
                                         "message": "Error en el formulario"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': f'Error al guardar los datos. {ex}'})

        elif action == 'deletetareaslimp':
            try:
                id = int(encrypt(request.POST['id']))
                hd_grupos = HdMantenimientoGruCategoria.objects.get(id=id)
                hd_grupos.status = False
                hd_grupos.save(request)
                log(u'Eliminó el registro: %s - %s'%(hd_grupos.__str__(),hd_grupos.pk),request,'delete')
                return JsonResponse({'error':False,'mensaje':u'Registro eliminado'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'error': True, 'mensaje': f'Error al guardar los datos. {ex}'})

        elif action == 'actdestarea':
            try:
                id = int(encrypt(request.POST['id']))
                hd_grupos = HdMantenimientoGruCategoria.objects.get(id=id)
                hd_grupos.activo = not hd_grupos.activo
                hd_grupos.save(request)
                log(u'Cambio de estado a la tarea: %s - %s - %s'%(hd_grupos.__str__(),hd_grupos.activo,hd_grupos.pk),request,'delete')
                return JsonResponse({'error':False,'mensaje':u'Registro eliminado'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'error': True, 'mensaje': f'Error al guardar los datos. {ex}'})
        elif action == 'adddanios':
            try:
                form = HdMantenimientoGruDaniosForm(request.POST)
                if form.is_valid():
                    hd_grupos = HdMantenimientoGruDanios(
                        grupocategoria= form.cleaned_data['tipobien'],
                        descripcion = form.cleaned_data['descripcion']
                    )
                    hd_grupos.save(request)
                    log(u'Agregó un mantenimiento de daño: %s'%(hd_grupos.__str__()),request,'add')
                    return JsonResponse({'result':False,'message':'Registro exitosa'})
                else:
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()],
                                         "message": "Error en el formulario"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': f'Error al guardar los datos. {ex}'})

        elif action == 'editdanios':
            try:
                form = HdMantenimientoGruDaniosForm(request.POST)
                filtro= HdMantenimientoGruDanios.objects.get(id = int(encrypt(request.POST['id'])))
                if form.is_valid():
                    filtro.grupocategoria = form.cleaned_data['tipobien']
                    filtro.descripcion = form.cleaned_data['descripcion']
                    filtro.save(request)
                    log(u'Editó un mantenimiento de daño: %s'%(filtro.__str__()),request,'edit')
                    return JsonResponse({'result':False,'message':'Edicion exitosa'})
                else:
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()],
                                         "message": "Error en el formulario"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': f'Error al guardar los datos. {ex}'})

        elif action == 'deletedanios':
            try:
                id = int(encrypt(request.POST['id']))
                hd_grupos = HdMantenimientoGruDanios.objects.get(id=id)
                hd_grupos.status = False
                hd_grupos.save(request)
                log(u'Eliminó el registro: %s - %s'%(hd_grupos.__str__(),hd_grupos.pk),request,'delete')
                return JsonResponse({'error':False,'mensaje':u'Registro eliminado'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'error': True, 'mensaje': f'Error al guardar los datos. {ex}'})

        elif action == 'actdesdanios':
            try:
                id = int(encrypt(request.POST['id']))
                hd_grupos = HdMantenimientoGruDanios.objects.get(id=id)
                hd_grupos.activo = not hd_grupos.activo
                hd_grupos.save(request)
                log(u'Cambio de estado a la tareas de daños: %s - %s - %s'%(hd_grupos.__str__(),hd_grupos.activo,hd_grupos.pk),request,'delete')
                return JsonResponse({'error':False,'mensaje':u'Registro eliminado'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'error': True, 'mensaje': f'Error al guardar los datos. {ex}'})

        elif action == 'addresponsable':
            try:

                form = ResponsableActivoTraspasoForm(request.POST)
                if form.is_valid():
                    responsable = form.cleaned_data['persona']
                    activo = int(encrypt(request.POST['id']))
                    verificarsolicitud = SolicitudActivos.objects.filter(status=True, activo_id=activo,
                                                                         responsableasignacion_id=responsable,
                                                                         estado=1)
                    verificarresponsable = ActivoTecnologico.objects.filter(status=True, activotecnologico__responsable_id=responsable,
                                                                     pk=activo)
                    if not verificarsolicitud:
                        if not verificarresponsable:
                            solicitud = SolicitudActivos(activo_id=activo, responsableasignacion_id=responsable,
                                                         solicitante=persona,
                                                         fechasolicitud=datetime.now().date(),
                                                         accionorigen=2)
                            solicitud.save(request)
                            log(u'Envio de solicitud traspaso: %s' % solicitud, request, "add")
                            return JsonResponse({"result": 'ok','mensaje':'Registro de traspaso exitoso'}, safe=False)
                        return JsonResponse({"result": 'bad', "mensaje": u"El activo ya cuenta con el mismo responsable"},
                                            safe=False)
                    return JsonResponse({"result": 'bad', "mensaje": u"El activo mantiene una solicitud pendiente"},
                                        safe=False)
                else:
                    return JsonResponse({'result': 'bad', "form2": [{k: v[0]} for k, v in form.errors.items()],
                                         "mensaje": u"Error en el formulario"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result':'bad','mensaje':u'Error al procesar los datos'})

        elif action == 'addpiezaparte':
            try:
                form = HdPiezaPartesForm(request.POST, request.FILES)
                newfile = None
                if form.is_valid():
                    if 'imagen' in request.FILES:
                        newfile = request.FILES['imagen']
                        extension = newfile._name.split('.')
                        tam = len(extension)
                        exte = extension[tam - 1]
                        if newfile.size > 2194304:
                            return JsonResponse(
                                {"result": True, "mensaje": u"Error, el tamaño del archivo es mayor a 4 Mb."})
                        if exte in ['jpg', 'jpeg', 'png']:
                            newfile._name = generar_nombre("imagen", newfile._name)
                        else:
                            return JsonResponse(
                                {"result": True, "mensaje": u"Error, solo archivos .pdf, .jpg, .jpeg"})
                    piezaparte = HdPiezaPartes(descripcion=form.cleaned_data['descripcion'],
                                               estado=form.cleaned_data['estado'],
                                               imagen=newfile,
                                               grupocategoria=form.cleaned_data['tipobien'])
                    piezaparte.save(request)
                    log(u'Adicionó pieza parte: %s - %s' % (persona, piezaparte), request, 'add')
                    return JsonResponse({'result':False, 'message': u'Registro de pieza/parte exitoso'})
                else:
                    return JsonResponse({"result": False, "mensaje": f"Error al validar los datos.", "form": [{k: v[0]} for k, v in f.errors.items()]})
            except Exception as ex:
                pass

        elif action == 'deletepiezaparte':
            try:
                piezaparte = HdPiezaPartes.objects.get(pk=int(encrypt(request.POST['id'])))
                piezaparte.status = False
                piezaparte.save(request)
                messages.success(request, 'Se eliminó exitosamente.')
                log(u'Eliminó pieza/parte %s' % piezaparte, request, "del")
                return JsonResponse({"result": False,"error":False}, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True,"error":True, "mensaje": "Intentelo más tarde."}, safe=False)

        elif action == 'editpiezaparte':
            try:
                form = HdPiezaPartesForm(request.POST, request.FILES)
                newfile = None
                if form.is_valid():
                    if 'imagen' in request.FILES:
                        newfile = request.FILES['imagen']
                        extension = newfile._name.split('.')
                        tam = len(extension)
                        exte = extension[tam - 1]
                        if newfile.size > 2194304:
                            return JsonResponse(
                                {"result": True, "mensaje": u"Error, el tamaño del archivo es mayor a 4 Mb."})
                        if exte in ['jpg', 'jpeg', 'png']:
                            newfile._name = generar_nombre("imagen", newfile._name)
                        else:
                            return JsonResponse(
                                {"result": True, "mensaje": u"Error, solo archivos .pdf, .jpg, .jpeg"})
                    piezaparte = HdPiezaPartes.objects.get(id=int(encrypt(request.POST['id'])))
                    if not newfile:
                        newfile = piezaparte.imagen
                    piezaparte.descripcion = form.cleaned_data['descripcion']
                    piezaparte.estado = form.cleaned_data['estado']
                    piezaparte.imagen = newfile
                    piezaparte.grupocategoria = form.cleaned_data['tipobien']
                    piezaparte.save(request)
                    log(u'Editó pieza parte: %s - %s' % (persona, piezaparte), request, 'add')
                    return JsonResponse({'result': False, 'message': u'Registro de pieza/parte exitoso'})
            except Exception as ex:
                pass

        return HttpResponseRedirect(request.path)

    else:
        if 'action' in request.GET:
            data['action'] = action = request.GET['action']
            if action == 'addgruposistemas':
                try:
                    data['title'] = u'Adicionar Grupo Sistemas/Equipo'
                    form = HdGrupoSistemaEquipoForm()
                    data['perfilprin'] = perfilprin = request.session['perfilprincipal']
                    data['form'] = form
                    return render(request, "helpdesk_hdplanificacion/addgruposistemas.html", data)
                except Exception as ex:
                    pass
            if action == 'editgruposistemas':
                try:
                    data['title'] = u'Editar Grupo de Sistemas/Equipos'
                    data['grupo'] = grupo = HdGrupoSistemaEquipo.objects.get(pk=int(request.GET['id']))
                    data['perfilprin'] = perfilprin = request.session['perfilprincipal']
                    form = HdGrupoSistemaEquipoForm(initial={'descripcion': grupo.descripcion,
                                                             })
                    data['form'] = form
                    return render(request, "helpdesk_hdplanificacion/editgruposistemas.html", data)
                except Exception as ex:
                    pass
            if action == 'delgruposistemas':
                try:
                    data['title'] = u'Eliminar  Grupo de Sistemas/Equipos'
                    data['grupo'] = HdGrupoSistemaEquipo.objects.get(pk=int(request.GET['id']))
                    return render(request, "helpdesk_hdplanificacion/deletegruposistemas.html", data)
                except Exception as ex:
                    pass
            if action == 'viewgruposistemas':
                data['title'] = u'Registro de  Grupo de Sistemas/Equipos'
                search = None
                ids = None
                request.session['viewactivo'] = 2
                if 's' in request.GET:
                    search = request.GET['s']
                    ss = search.split(' ')
                    if len(ss) == 1:
                        gruposistemas = HdGrupoSistemaEquipo.objects.filter(descripcion__icontains=search, status=True)
                    elif len(ss) == 2:
                        gruposistemas = HdGrupoSistemaEquipo.objects.filter(Q(descripcion__icontains=ss[0])| Q(descripcion__icontains=ss[1]),Q(status=True))
                    else:
                        gruposistemas = HdGrupoSistemaEquipo.objects.filter(Q(descripcion__icontains=ss[0])| Q(descripcion__icontains=ss[1]),Q(status=True) )
                elif 'id' in request.GET:
                    ids = request.GET['id']
                    gruposistemas = HdGrupoSistemaEquipo.objects.filter(id=ids, status=True, persona_id=persona.id).order_by(
                        'estado_id')
                else:
                    gruposistemas = HdGrupoSistemaEquipo.objects.filter(status=True)
                paging = MiPaginador(gruposistemas, 10)
                p = 1
                try:
                    paginasesion = 1
                    if 'paginador' in request.session:
                        paginasesion = int(request.session['paginador'])
                    if 'page' in request.GET:
                        p = int(request.GET['page'])
                    else:
                        p = paginasesion
                    try:
                        page = paging.page(p)
                    except:
                        p = 1
                    page = paging.page(p)
                except:
                    page = paging.page(p)
                request.session['paginador'] = p
                data['paging'] = paging
                data['rangospaging'] = paging.rangos_paginado(p)
                data['page'] = page
                data['search'] = search if search else ""
                data['ids'] = ids if ids else ""
                data['grupo'] = page.object_list
                data['administrativo'] = persona
                return render(request, "helpdesk_hdplanificacion/viewgruposistemas.html", data)

            if action == 'addbien':
                try:
                    data['title'] = u'Adicionar Bienes'
                    form = HdBienForm()
                    data['perfilprin'] = perfilprin = request.session['perfilprincipal']
                    data['form'] = form
                    return render(request, "helpdesk_hdplanificacion/addbien.html", data)
                except Exception as ex:
                    pass
            if action == 'editbien':
                try:
                    data['title'] = u'Editar Bien'
                    data['bien'] = bien = HdBien.objects.get(pk=int(request.GET['id']))
                    data['perfilprin'] = perfilprin = request.session['perfilprincipal']
                    form = HdBienForm(initial={'cantidad': bien.cantidad,
                                               'sistemaequipo': bien.sistemaequipo,
                                               'observacion': bien.observacion,
                                               'bloque': bien.ubicacion.bloque if bien.ubicacion.bloque else None,
                                               'ubicacion': bien.ubicacion if bien.ubicacion else None,
                                                             })
                    if bien.gruposistema:
                        form.editar(bien.gruposistema)
                    data['form'] = form
                    return render(request, "helpdesk_hdplanificacion/editbien.html", data)
                except Exception as ex:
                    pass

            elif action == 'editmaterial':
                try:
                    data['title'] = u'Editar Material'
                    data['mantmaterial'] = mantmaterial = HdMaterialMantenimiento.objects.get(pk=int(request.GET['id']))
                    grupo = None
                    agente = None
                    ayudantes = None

                    ayudantes = HdDetalle_Grupo.objects.filter(pk__in=HdMaterialMantenimiento_Responsable.objects.values_list('agente__id').filter(materialmantenimiento=mantmaterial.id, status=True), status=True)
                        # if incidente.ultimo_registro().estadoasignacion == 1:
                        #     if incidente.ultimo_registro().grupo:
                        #         grupo = incidente.ultimo_registro().grupo
                        #     if incidente.ultimo_registro().agente:
                        #         agente = incidente.ultimo_registro().agente
                        # else:
                        #     agente = HdDetalle_Grupo.objects.filter(persona=persona, grupo__tipoincidente=incidente.tipoincidente)[0]
                        #     grupo = agente.grupo

                    form = HdMaterialMantenimientoForm(initial={
                                                     'tipomantenimiento': mantmaterial.tipomantenimiento,
                                                     'bloque': mantmaterial.ubicacion.bloque if mantmaterial.ubicacion.bloque else None,
                                                     'ubicacion': mantmaterial.ubicacion if mantmaterial.ubicacion else None,
                                                     'tipobien': mantmaterial.tipobien,
                                                     'proceso': mantmaterial.proceso,
                                                     'ayudantes': ayudantes
                                                     })

                    if mantmaterial.gruposistema:
                        form.editar(mantmaterial.gruposistema)
                    if mantmaterial.bien:
                        form.editarbien(mantmaterial.bien)



                    form.resolver()

                    if grupo:
                        if grupo.tipoincidente.id == 2:
                            es_tics = True







                    data['mantmateriales'] = HdMaterialMantenimiento_Material.objects.filter(materialmantenimiento=int(mantmaterial.pk),status=True)
                    data['formmateriales'] = HdMaterialMantenimiento_MaterialForm()
                    data['form'] = form
                    return render(request, "helpdesk_hdplanificacion/editmaterial.html", data)
                except Exception as ex:
                    pass
            if action == 'delbien':
                try:
                    data['title'] = u'Eliminar  Bien'
                    data['bien'] = HdBien.objects.get(pk=int(request.GET['id']))
                    return render(request, "helpdesk_hdplanificacion/deletebien.html", data)
                except Exception as ex:
                    pass
            if action == 'viewbien':
                data['title'] = u'Registro de  Bien'
                search = None
                ids = None
                request.session['viewactivo'] = 3
                if 's' in request.GET:
                    search = request.GET['s']
                    ss = search.split(' ')
                    if len(ss) == 1:
                        bien = HdBien.objects.filter(ubicacion__bloque__nombre__icontains=search,
                                                     gruposistema__descripcion__icontains=search,sistemaequipo=search,
                                                      status=True,observacion=search)
                    elif len(ss) == 2:
                        bien = HdBien.objects.filter(
                                                     Q(ubicacion__bloque__nombre__icontains=ss[0])|
                                                     Q(gruposistema__descripcion__icontains=ss[0])| Q(sistemaequipo=ss[0])|Q(observacion=ss[0]), status=True)

                    else:
                        bien = HdBien.objects.filter(Q(ubicacion__bloque__nombre__icontains=ss[1]),
                                                                            Q(descripcion__icontains=ss[1])| Q(ubicacion__bloque__nombre__icontains=ss[0])|Q(gruposistema__descripcion__icontains=ss[0])|
                                                     Q(sistemaequipo=ss[0]),
                                                                            Q(status=True))
                elif 'id' in request.GET:
                    ids = request.GET['id']
                    bien = HdBien.objects.filter(id=ids, status=True)
                else:
                    bien = HdBien.objects.filter(status=True)
                paging = MiPaginador(bien, 10)
                p = 1
                try:
                    paginasesion = 1
                    if 'paginador' in request.session:
                        paginasesion = int(request.session['paginador'])
                    if 'page' in request.GET:
                        p = int(request.GET['page'])
                    else:
                        p = paginasesion
                    try:
                        page = paging.page(p)
                    except:
                        p = 1
                    page = paging.page(p)
                except:
                    page = paging.page(p)
                request.session['paginador'] = p
                data['paging'] = paging
                data['rangospaging'] = paging.rangos_paginado(p)
                data['page'] = page
                data['search'] = search if search else ""
                data['ids'] = ids if ids else ""
                data['bien'] = page.object_list
                data['administrativo'] = persona
                data['fecha'] = datetime.now().date()
                data['grupo'] = HdGrupoSistemaEquipo.objects.filter(status=True)
                return render(request, "helpdesk_hdplanificacion/viewbien.html", data)



            if action == 'addconf':
                try:
                    data['title'] = u'Adicionar Configuración Frecuencia'
                    form = HdConfFrecuenciaForm()
                    data['perfilprin'] = perfilprin = request.session['perfilprincipal']
                    data['form'] = form
                    return render(request, "helpdesk_hdplanificacion/addconf.html", data)
                except Exception as ex:
                    pass
            if action == 'editconf':
                try:
                    data['title'] = u'Editar  Configuración Frecuencia'
                    data['configuracion'] = configuracion = HdConfFrecuencia.objects.get(pk=int(request.GET['id']))
                    data['perfilprin'] = perfilprin = request.session['perfilprincipal']
                    form = HdConfFrecuenciaForm(initial={'duracion': configuracion.duracion,
                                                         'cantidad': configuracion.cantidad,
                                                             })
                    data['form'] = form
                    return render(request, "helpdesk_hdplanificacion/editbien.html", data)
                except Exception as ex:
                    pass
            if action == 'delconf':
                try:
                    data['title'] = u'Eliminar   Configuración Frecuencia'
                    data['configuracion'] = HdConfFrecuencia.objects.get(pk=int(request.GET['id']))
                    return render(request, "helpdesk_hdplanificacion/deleteconf.html", data)
                except Exception as ex:
                    pass
            if action == 'viewconf':
                data['title'] = u'Registro de   Configuración Frecuencia'
                search = None
                request.session['viewactivo'] = 4
                ids = None
                if 's' in request.GET:
                    search = request.GET['s']
                    ss = search.split(' ')
                    if len(ss) == 1:
                        conf = HdConfFrecuencia.objects.filter(duracion__icontains=search,cantidad__icontains=search, status=True)
                    elif len(ss) == 2:
                        conf = HdConfFrecuencia.objects.filter(Q(duracion__icontains=ss[0])|
                                                     Q(cantidad__icontains=ss[0]),
                                                     status=True)

                    else:
                        conf = HdConfFrecuencia.objects.filter(Q(duracion__icontains=ss[0])|Q(cantidad__icontains=ss[1]),
                                                                            Q(status=True))
                elif 'id' in request.GET:
                    ids = request.GET['id']
                    conf = HdConfFrecuencia.objects.filter(id=ids, status=True)
                else:
                    conf = HdConfFrecuencia.objects.filter(status=True)
                paging = MiPaginador(conf, 10)
                p = 1
                try:
                    paginasesion = 1
                    if 'paginador' in request.session:
                        paginasesion = int(request.session['paginador'])
                    if 'page' in request.GET:
                        p = int(request.GET['page'])
                    else:
                        p = paginasesion
                    try:
                        page = paging.page(p)
                    except:
                        p = 1
                    page = paging.page(p)
                except:
                    page = paging.page(p)
                request.session['paginador'] = p
                data['paging'] = paging
                data['rangospaging'] = paging.rangos_paginado(p)
                data['page'] = page
                data['search'] = search if search else ""
                data['ids'] = ids if ids else ""
                data['configuracion'] = page.object_list
                data['administrativo'] = persona
                return render(request, "helpdesk_hdplanificacion/viewconf.html", data)

            if action == 'addfrecuencia':
                try:
                    data['title'] = u'Adicionar  Frecuencia'
                    form = HdFrecuenciaForm()
                    data['perfilprin'] = perfilprin = request.session['perfilprincipal']
                    data['form'] = form
                    return render(request, "helpdesk_hdplanificacion/addfrecuencia.html", data)
                except Exception as ex:
                    pass
            if action == 'editfrecuencia':
                try:
                    data['title'] = u'Editar   Frecuencia'
                    data['frecuencia'] = frecuencia = HdFrecuencia.objects.get(pk=int(request.GET['id']))
                    data['perfilprin'] = perfilprin = request.session['perfilprincipal']
                    form = HdFrecuenciaForm(initial={'tipomantenimiento': frecuencia.tipomantenimiento,
                                                         'descripcion': frecuencia.descripcion,
                                                         'consideracion': frecuencia.consideracion,
                                                         'proceso': frecuencia.proceso,
                                                         })
                    if frecuencia.gruposistema:
                        form.editar(frecuencia.gruposistema)
                    if frecuencia.bien:
                        form.editarbien(frecuencia.bien)
                    if frecuencia.frecuencia:
                        form.editarfrecuencia(frecuencia.frecuencia)
                    data['form'] = form
                    return render(request, "helpdesk_hdplanificacion/editfrecuencia.html", data)
                except Exception as ex:
                    pass
            if action == 'delfrecuencia':
                try:
                    data['title'] = u'Eliminar   Frecuencia'
                    data['frecuencia'] = HdFrecuencia.objects.get(pk=int(request.GET['id']))
                    return render(request, "helpdesk_hdplanificacion/deletefrecuencia.html", data)
                except Exception as ex:
                    pass
            if action == 'viewfrecuencia':
                data['title'] = u'Registro  Frecuencia'
                search = None
                request.session['viewactivo'] = 5
                ids = None
                if 's' in request.GET:
                    search = request.GET['s']
                    ss = search.split(' ')
                    # if len(ss) == 1:
                    #     frecuencia = HdFrecuencia.objects.filter(descripcion__icontains=search, consideracion__icontains=search,
                    #                                            status=True,bien__sistemaequipo__icontains=search,gruposistema__descripcion__icontains=search,
                    #                                              proceso__icontains=search,frecuencia__duracion__icontains=search)
                    if len(ss) == 1:
                        frecuencia = HdFrecuencia.objects.filter(Q(descripcion__icontains=ss[0])|
                                                               Q(consideracion__icontains=ss[0])|Q(bien__sistemaequipo__icontains=ss[0])|Q(gruposistema__descripcion__icontains=ss[0]),
                                                               status=True)

                    else:
                        frecuencia = HdFrecuencia.objects.filter(Q(descripcion__icontains=ss[0])|
                                                               Q(consideracion__icontains=ss[0])|Q(bien__sistemaequipo__icontains=ss[0])|Q(gruposistema__descripcion__icontains=ss[0])|Q(descripcion__icontains=ss[1])|
                                                               Q(consideracion__icontains=ss[1])|Q(bien__sistemaequipo__icontains=ss[1])|Q(gruposistema__descripcion__icontains=ss[1]),
                                                               Q(status=True))


                elif 'id' in request.GET:
                    ids = request.GET['id']
                    frecuencia = HdFrecuencia.objects.filter(id=ids, status=True)
                else:
                    frecuencia = HdFrecuencia.objects.filter(status=True)
                paging = MiPaginador(frecuencia, 10)
                p = 1
                try:
                    paginasesion = 1
                    if 'paginador' in request.session:
                        paginasesion = int(request.session['paginador'])
                    if 'page' in request.GET:
                        p = int(request.GET['page'])
                    else:
                        p = paginasesion
                    try:
                        page = paging.page(p)
                    except:
                        p = 1
                    page = paging.page(p)
                except:
                    page = paging.page(p)
                request.session['paginador'] = p
                data['paging'] = paging
                data['rangospaging'] = paging.rangos_paginado(p)
                data['page'] = page
                data['search'] = search if search else ""
                data['ids'] = ids if ids else ""
                data['frecuencia'] = page.object_list
                data['administrativo'] = persona
                data['fecha'] = datetime.now().date()
                data['grupo'] = HdGrupoSistemaEquipo.objects.filter(status=True)
                return render(request, "helpdesk_hdplanificacion/viewfrecuencia.html", data)

            if action == 'addmantenimiento':
                try:
                    data['title'] = u'Adicionar  Mantenimiento Mensual'
                    form = HdCronogramaMantenimientoForm()
                    data['form'] = form
                    return render(request, "helpdesk_hdplanificacion/addmantenimiento.html", data)
                except Exception as ex:
                    pass
            if action == 'editmantenimiento':
                try:
                    data['title'] = u'Editar   Mantenimiento Mensual'
                    data['mantenimiento'] = mantenimiento = HdCronogramaMantenimiento.objects.get(pk=int(request.GET['id']))
                    data['perfilprin'] = perfilprin = request.session['perfilprincipal']
                    form = HdCronogramaMantenimientoForm(initial={'tipomantenimiento': mantenimiento.tipomantenimiento,
                                                                  'desde': mantenimiento.desde,
                                                                  'hasta': mantenimiento.hasta})
                    if mantenimiento.gruposistema:
                        form.editar(mantenimiento.gruposistema)
                    if mantenimiento.proveedor:
                        form.editar_proveedor(mantenimiento.proveedor)
                    data['form'] = form
                    return render(request, "helpdesk_hdplanificacion/editmantenimiento.html", data)
                except Exception as ex:
                    pass
            if action == 'delmantenimiento':
                try:
                    data['title'] = u'Eliminar   Mantenimiento Mensual'
                    data['mantenimiento'] = HdCronogramaMantenimiento.objects.get(pk=int(request.GET['id']))
                    return render(request, "helpdesk_hdplanificacion/deletemantenimiento.html", data)
                except Exception as ex:
                    pass

            if action == 'addreparacion':
                try:
                    data['title'] = u'Adicionar  Reparación'
                    form = HdReparacionForm()
                    data['perfilprin'] = perfilprin = request.session['perfilprincipal']
                    data['form'] = form
                    return render(request, "helpdesk_hdplanificacion/addreparacion.html", data)
                except Exception as ex:
                    pass
            if action == 'editreparacion':
                try:
                    data['title'] = u'Editar Reparación'
                    data['reparacion'] = reparacion = HdReparacion.objects.get(pk=int(request.GET['id']))
                    data['perfilprin'] = perfilprin = request.session['perfilprincipal']
                    form = HdReparacionForm(initial={
                                                     'bloque': reparacion.ubicacion.bloque if reparacion.ubicacion.bloque else None,
                                                     'ubicacion': reparacion.ubicacion if reparacion.ubicacion else None,
                                                         })
                    if reparacion.gruposistema:
                        form.editar(reparacion.gruposistema)

                    data['form'] = form
                    return render(request, "helpdesk_hdplanificacion/editreparacion.html", data)
                except Exception as ex:
                    pass
            if action == 'delreparacion':
                try:
                    data['title'] = u'Eliminar Reparación'
                    data['reparacion'] = HdReparacion.objects.get(pk=int(request.GET['id']))
                    return render(request, "helpdesk_hdplanificacion/deletereparacion.html", data)
                except Exception as ex:
                    pass
            if action == 'viewreparacion':
                try:
                    data['title'] = u'Reparación'
                    search = None
                    ids = None
                    if 's' in request.GET:
                        search = request.GET['s']
                        ss = search.split(' ')

                        if len(ss) == 1:
                            reparacion = HdReparacion.objects.filter(
                                Q(gruposistema__descripcion__icontains=ss[0])| Q(ubicacion__bloque__nombre__icontains=ss[0] )
                                                | Q(ubicacion__ubicacion__nombre__icontains=ss[0])
                                ,Q(status=True)).order_by(
                                'fecha_creacion')
                        else:
                            reparacion = HdReparacion.objects.filter(
                                Q(gruposistema__descripcion__icontains=ss[0]) | Q(
                                    ubicacion__bloque__nombre__icontains=ss[0]) | Q(
                                    gruposistema__descripcion__icontains=ss[1]) |
                                Q(ubicacion__bloque__nombre__icontains=ss[1]), Q(status=True)).order_by(
                                'fecha_creacion')
                    elif 'id' in request.GET:
                        ids = request.GET['id']
                        reparacion = HdReparacion.objects.filter(id=ids, status=True).order_by(
                            'fecha_creacion')
                    else:
                        reparacion = HdReparacion.objects.filter(status=True).order_by('fecha_creacion')
                    paging = MiPaginador(reparacion, 10)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['reparacion'] = page.object_list
                    data['administrativo'] = persona
                    return render(request, "helpdesk_hdplanificacion/viewreparacion.html", data)

                except Exception as ex:
                    pass


            if action == 'adddetreparacion':
                try:
                    data['title'] = u'Adicionar  Detalle Reparación'
                    form = HdDetalle_ReparacionForm()
                    data['repa'] = crono = HdReparacion.objects.get(pk=int(request.GET['reparacion']))
                    data['form'] = form
                    return render(request, "helpdesk_hdplanificacion/adddetreparacion.html", data)
                except Exception as ex:
                    pass
            if action == 'editdetreparacion':
                try:
                    data['title'] = u'Editar   Detalle Reparación'
                    data['detreparacion'] = reparacion = HdDetalle_Reparacion.objects.get(pk=int(request.GET['id']))
                    data['perfilprin'] = perfilprin = request.session['perfilprincipal']
                    form = HdDetalle_ReparacionForm(initial={
                                                     'descripcion': reparacion.descripcion,
                                                     'cantidad': reparacion.cantidad,
                                                     })
                    if reparacion.bien:
                        form.editarbien(reparacion.bien)

                    data['form'] = form
                    return render(request, "helpdesk_hdplanificacion/editdetreparacion.html", data)
                except Exception as ex:
                    pass
            if action == 'deldetreparacion':
                try:
                    data['title'] = u'Eliminar   Detalle Reparación'
                    data['detreparacion'] = HdDetalle_Reparacion.objects.get(pk=int(request.GET['id']))
                    return render(request, "helpdesk_hdplanificacion/deletedetreparacion.html", data)
                except Exception as ex:
                    pass
            if action == 'viewdetreparacion':
                try:
                    data['title'] = 'Registro de Detalle de Reparación'
                    detreparacion=None
                    search = None
                    ids = None
                    if 'id' in request.GET:
                        data['repa']  = HdReparacion.objects.get(pk=int(request.GET['id'])).pk
                    if 'reparacion' in request.GET:
                        data['repa']  = HdReparacion.objects.get(pk=int(request.GET['reparacion'])).pk

                    if 's' in request.GET:
                        search = request.GET['s']
                        ss = search.split(' ')
                        # if len(ss) == 1:
                        #     detreparacion = HdDetalle_Reparacion.objects.filter(bien__sistemaequipo__icontains=search, descripcion__icontains=search,
                        #                                            status=True,cantidad__icontains=search,reparacion=int(request.GET['id']))


                        if len(ss) == 1:
                            detreparacion = HdDetalle_Reparacion.objects.filter(Q(bien__sistemaequipo__icontains=ss[0])|
                                                                   Q(descripcion__icontains=ss[0])|Q(cantidad__icontains=ss[0]),
                                                                   status=True,reparacion=int(request.GET['id']))

                        else:
                            detreparacion = HdDetalle_Reparacion.objects.filter(Q(bien__sistemaequipo__icontains=ss[0])|Q(bien__sistemaequipo__icontains=ss[1])|
                                                                   Q(descripcion__icontains=ss[0])|Q(cantidad__icontains=ss[0]),
                                                                   Q(descripcion__icontains=ss[1])|Q(cantidad__icontains=ss[1]),
                                                                   Q(status=True),reparacion=int(request.GET['id']))
                    else:

                        if 'id' in request.GET:
                            detreparacion = HdDetalle_Reparacion.objects.filter(status=True, reparacion=int(request.GET['id']))
                        if 'reparacion' in request.GET:
                            detreparacion = HdDetalle_Reparacion.objects.filter(status=True,
                                                                   reparacion=int(request.GET['reparacion']))

                    paging = MiPaginador(detreparacion, 25)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['page'] = page
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['detreparacion'] = page.object_list
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    return render(request, "helpdesk_hdplanificacion/viewdetreparacion.html", data)
                except Exception as ex:
                    pass

            if action == 'addmaterial':
                try:
                    data['title'] = u'Adicionar Mantenimiento  Material '
                    form = HdMaterialMantenimientoForm()
                    # data['repa'] = crono = HdMaterialMantenimiento.objects.get(pk=int(request.GET['reparacion']))
                    # data['materialesincidentes'] = incidente.hdmaterial_incidente_set.filter(status=True)
                    data['formmateriales'] = HdMaterialMantenimiento_MaterialForm()
                    data['form'] = form
                    return render(request, "helpdesk_hdplanificacion/addmantmaterial.html", data)
                except Exception as ex:
                    pass
            if action == 'editmateriales':
                try:
                    data['title'] = u'Editar   Detalle Reparación'
                    data['detreparacion'] = reparacion = HdDetalle_Reparacion.objects.get(pk=int(request.GET['id']))
                    data['perfilprin'] = perfilprin = request.session['perfilprincipal']
                    form = HdDetalle_ReparacionForm(initial={
                        'descripcion': reparacion.descripcion,
                        'cantidad': reparacion.cantidad,
                    })
                    if reparacion.bien:
                        form.editarbien(reparacion.bien)

                    data['form'] = form
                    return render(request, "helpdesk_hdplanificacion/editdetreparacion.html", data)
                except Exception as ex:
                    pass
            if action == 'delmateriales':
                try:
                    data['title'] = u'Eliminar   Detalle Reparación'
                    data['detreparacion'] = HdDetalle_Reparacion.objects.get(pk=int(request.GET['id']))
                    return render(request, "helpdesk_hdplanificacion/deletedetreparacion.html", data)
                except Exception as ex:
                    pass
            if action == 'viewmateriales':
                try:
                    data['title'] = u'Registro Materiales'
                    search = None
                    ids = None
                    if 's' in request.GET:
                        search = request.GET['s']
                        ss = search.split(' ')
                        if len(ss) == 1:
                            material = HdMaterialMantenimiento.objects.filter(Q(gruposistema__descripcion__icontains=ss[0]) | Q(ubicacion__bloque__nombre__icontains=ss[0])| Q(bien__sistemaequipo__icontains=ss[0]),Q(status=True)).order_by('fecha_creacion')
                        else:
                            material = HdMaterialMantenimiento.objects.filter(Q(gruposistema__descripcion__icontains=ss[0]) | Q(ubicacion__bloque__nombre__icontains=ss[0]) | Q(bien__sistemaequipo__icontains=ss[0]) |Q(gruposistema__descripcion__icontains=ss[1]) | Q(ubicacion__bloque__nombre__icontains=ss[1]) | Q(bien__sistemaequipo__icontains=ss[1]), Q(status=True)).order_by('fecha_creacion')
                    elif 'id' in request.GET:
                        ids = request.GET['id']
                        material = HdMaterialMantenimiento.objects.filter(id=ids, status=True).order_by('fecha_creacion')
                    else:
                        material = HdMaterialMantenimiento.objects.filter(status=True).order_by('fecha_creacion')
                    paging = MiPaginador(material, 10)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['material'] = page.object_list
                    data['administrativo'] = persona
                    return render(request, "helpdesk_hdplanificacion/viewmaterial.html", data)

                except Exception as ex:
                    pass


            if action == 'delmantmaterial':
                try:
                    data['title'] = u'Eliminar Mantenimiemto Material'
                    data['mantmaterial'] = HdMaterialMantenimiento.objects.get(pk=int(request.GET['id']))
                    return render(request, "helpdesk_hdplanificacion/deletemantenimientomat.html", data)
                except Exception as ex:
                    pass



            if action == 'adddetalle':
                try:
                    data['title'] = u'Adicionar  Detalle Mantenimiento'
                    form = HdDetCronogramaMantenimientoForm()
                    data['crono'] = crono = HdCronogramaMantenimiento.objects.get(pk=int(request.GET['cronograma']))
                    data['form'] = form
                    return render(request, "helpdesk_hdplanificacion/adddetalle.html", data)
                except Exception as ex:
                    pass

            if action == 'addmantenimientoactivo':
                try:
                    data['title'] = u'Adicionar  Detalle Mantenimiento'
                    form = HdMantenimientosActivosForm()
                    data['crono'] = crono = HdCronogramaMantenimiento.objects.get(pk=int(request.GET['cronograma']))
                    data['form'] = form
                    return render(request, "helpdesk_hdplanificacion/addmantenimientoactivo.html", data)
                except Exception as ex:
                    pass

            if action == 'editmantenimientoactivo':
                try:
                    data['title'] = u'Editar  Detalle Mantenimiento'
                    data['manpreventivos'] = mantenimiento = HdDetMantenimientosActivos.objects.get(status=True,pk = int(encrypt(request.GET['id'])))
                    data['grupocategoria'] = HdGruposCategoria.objects.filter(status=True)
                    data['tareasmantenimiento'] = HdMantenimientoGruCategoria.objects.filter(grupocategoria=mantenimiento.tipoactivoc,status=True, activo=True)
                    data['tareasactivo'] = HdTareasActivosPreventivos.objects.values_list('grupos_id', flat=True).filter(mantenimiento=mantenimiento, status=True)
                    data['piezaparteactivo'] = HdPiezaParteActivosPreventivos.objects.filter(mantenimiento=mantenimiento, status=True)
                    data['daniomantenimiento'] = HdMantenimientoGruDanios.objects.filter(
                        grupocategoria=mantenimiento.tipoactivoc, status=True, activo=True)
                    data['danioactivo'] = HdTareasActivosPreventivosDanios.objects.filter(
                        mantenimiento=mantenimiento, status=True)
                    data['tipo'] = TIPO_MANTENIMIENTO
                    form = HdMantenimientosActivosForm(initial={'activotecno': mantenimiento.activotecno,
                                                                'tipoactivo': mantenimiento.tipoactivoc,
                                                                'persona':mantenimiento.activotecno.activotecnologico.responsable,
                                                                'estusu': mantenimiento.estusu,
                                                                'archivo': mantenimiento.archivo,
                                                                'fecha': mantenimiento.fecha,
                                                                'horamax': mantenimiento.horamax,
                                                                'minutomax': mantenimiento.minutomax,
                                                                'estfrec': mantenimiento.funcionarecibe,
                                                                'estfent': mantenimiento.funcionaentrega,
                                                                'marca': mantenimiento.marca,
                                                                'modelo': mantenimiento.modelo,
                                                                # 'caracteristica': mantenimiento.caracteristicas,
                                                                'bsugiere': mantenimiento.sbequipo,
                                                                'dsugiere': mantenimiento.descbaja,
                                                                # 'piezaparte': piezaparte,
                                                                'observacion': mantenimiento.observaciones})
                    form.cargar_mantenimiento(mantenimiento)
                    data['form'] = form
                    return render(request, "helpdesk_hdplanificacion/editmantenimientoactivo.html", data)
                except Exception as ex:
                    pass

            if action == 'editdetalle':
                try:
                    data['title'] = u'Editar   Detalle Mantenimiento'
                    data['detalle'] = detalle = HdDetCronogramaMantenimiento.objects.get(pk=int(request.GET['id']))
                    data['perfilprin'] = perfilprin = request.session['perfilprincipal']
                    form = HdDetCronogramaMantenimientoForm(initial={'inventario': detalle.inventario,
                                                     'mes': detalle.mes,
                                                     'descripcion': detalle.descripcion,
                                                     'cantidad': detalle.cantidad,
                                                     'bloque': detalle.ubicacion.bloque if detalle.ubicacion.bloque else None,
                                                     'ubicacion': detalle.ubicacion if detalle.ubicacion else None,
                                                     })
                    if detalle.bien:
                        form.editarbien(detalle.bien)

                    data['form'] = form
                    return render(request, "helpdesk_hdplanificacion/editdetalle.html", data)
                except Exception as ex:
                    pass
            if action == 'deldetalle':
                try:
                    data['title'] = u'Eliminar   Detalle Mantenimiento'
                    data['detalle'] = HdDetCronogramaMantenimiento.objects.get(pk=int(request.GET['id']))
                    return render(request, "helpdesk_hdplanificacion/deletedetalle.html", data)
                except Exception as ex:
                    pass
            if action == 'viewdetalle':
                try:
                    data['title'] = 'Registro de Detalle de Mantenimientos'
                    detalle=None
                    search = None
                    ids = None
                    if 'id' in request.GET:
                        data['crono']  = HdCronogramaMantenimiento.objects.get(pk=int(request.GET['id'])).pk
                    if 'cronograma' in request.GET:
                        data['crono']  = HdCronogramaMantenimiento.objects.get(pk=int(request.GET['cronograma'])).pk

                    if 's' in request.GET:
                        search = request.GET['s']
                        ss = search.split(' ')
                        if len(ss) == 1:
                            detalle = HdDetMantenimientosActivos.objects.filter((Q(activotecno__activotecnologico__codigogobierno=search) |
                                                                                 Q(activotecno__activotecnologico__codigointerno=search) |
                                                                                 Q(activotecno__activotecnologico__responsable__nombres__icontains=search)) & Q(cronograma=int(request.GET['id'])))
                        elif len(ss) == 2:
                            detalle = HdDetMantenimientosActivos.objects.filter((Q(activotecno__activotecnologico__responsable__apellido1__icontains=ss[0]) &
                                                                                 Q(activotecno__activotecnologico__responsable__apellido2__icontains=ss[1]) & Q(cronograma=int(request.GET['id']))))
                    else:
                        if 'id' in request.GET:
                            detalle = HdDetMantenimientosActivos.objects.filter(status=True, cronograma=int(request.GET['id']))
                        if 'cronograma' in request.GET:
                            detalle = HdDetMantenimientosActivos.objects.filter(status=True,
                                                                   cronograma=int(request.GET['cronograma']))

                    paging = MiPaginador(detalle, 25)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['page'] = page
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['detalle'] = page.object_list
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    return render(request, "helpdesk_hdplanificacion/viewdetalle.html", data)
                except Exception as ex:
                    pass

            if action == 'addreparacion':
                try:
                    data['title'] = u'Adicionar  Reparación'
                    form = HdReparacionForm()
                    data['perfilprin'] = perfilprin = request.session['perfilprincipal']
                    data['form'] = form
                    return render(request, "helpdesk_hdplanificacion/addreparacion.html", data)
                except Exception as ex:
                    pass
            if action == 'editreparacion':
                try:
                    data['title'] = u'Editar Reparación'
                    data['reparacion'] = reparacion = HdReparacion.objects.get(pk=int(request.GET['id']))
                    data['perfilprin'] = perfilprin = request.session['perfilprincipal']
                    form = HdReparacionForm(initial={
                        'bloque': reparacion.ubicacion.bloque if reparacion.ubicacion.bloque else None,
                        'ubicacion': reparacion.ubicacion if reparacion.ubicacion else None,
                    })
                    if reparacion.gruposistema:
                        form.editar(reparacion.gruposistema)

                    data['form'] = form
                    return render(request, "helpdesk_hdplanificacion/editreparacion.html", data)
                except Exception as ex:
                    pass
            if action == 'delreparacion':
                try:
                    data['title'] = u'Eliminar Reparación'
                    data['reparacion'] = HdReparacion.objects.get(pk=int(request.GET['id']))
                    return render(request, "helpdesk_hdplanificacion/deletereparacion.html", data)
                except Exception as ex:
                    pass
            if action == 'delcronogramasemanal':
                try:
                    data['title'] = u'Eliminar Cronograma Semanal'
                    data['crono'] = HdCronogramaMantenimientoSem.objects.get(pk=int(request.GET['id']))
                    return render(request, "helpdesk_hdplanificacion/deletecronogramasemanal.html", data)
                except Exception as ex:
                    pass
            if action == 'viewcronograma':
                try:
                    data['title'] = u'Cronograma Semanal'
                    search = None
                    ids = None
                    cronogramasem = HdCronogramaMantenimientoSem.objects.filter(status=True).order_by('fecha_creacion')
                    id_estado=0
                    idgrupo=0

                    if 's' in request.GET:
                        search = request.GET['s']
                        ss = search.split(' ')
                        if len(ss) == 1:
                            cronogramasem = HdCronogramaMantenimientoSem.objects.filter(
                                                                     gruposistema__descripcion__icontains=search,
                                                                     status=True).order_by('fecha_creacion')
                        elif len(ss) == 2:
                            cronogramasem = HdCronogramaMantenimientoSem.objects.filter(
                                Q(gruposistema__descripcion__icontains=ss[0]) | Q(gruposistema__descripcion__icontains=ss[1]), Q(status=True)).order_by('fecha_creacion')
                        else:
                            cronogramasem = HdCronogramaMantenimientoSem.objects.filter(
                                Q(gruposistema__descripcion__icontains=ss[0]) | Q(gruposistema__descripcion__icontains=ss[1])  , Q(status=True)).order_by('fecha_creacion')
                    elif 'id' in request.GET:
                        ids = request.GET['id']
                        cronogramasem = HdCronogramaMantenimientoSem.objects.filter(id=ids, status=True).order_by(
                            'fecha_creacion')

                    if 'idg' in request.GET and not 'id_estado' in request.GET:
                        idgrupo = int(request.GET['idg'])
                        if idgrupo != 0:
                            cronogramasem = HdCronogramaMantenimientoSem.objects.filter(status=True,
                                                                                  gruposistema__id=idgrupo).order_by(
                                'fecha_creacion')
                    elif 'id_estado' in request.GET  and not 'idg' in request.GET:
                        id_estado = int(request.GET['id_estado'])
                        if id_estado != 0:
                            cronogramasem = HdCronogramaMantenimientoSem.objects.filter(status=True,
                                                                                  mes=id_estado).order_by(
                                'fecha_creacion')
                    elif 'id_estado' in request.GET and 'idg' in request.GET:
                        id_estado = int(request.GET['id_estado'])
                        idgrupo = int(request.GET['idg'])
                        if id_estado != 0 and idgrupo != 0:
                            cronogramasem = HdCronogramaMantenimientoSem.objects.filter(status=True,
                                                                                  mes=id_estado,
                                                                                  gruposistema__id=idgrupo).order_by(
                                'fecha_creacion')

                    paging = MiPaginador(cronogramasem, 10)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['cronogramasem'] = page.object_list
                    data['cronogramas'] = HdDetCronogramaMantenimiento.objects.filter(status=True)
                    data['administrativo'] = persona
                    data['fecha'] = datetime.now().date()
                    data['grupo'] = HdGrupoSistemaEquipo.objects.filter(status=True)
                    data['bloque'] = HdBloqueUbicacion.objects.filter(status=True)
                    data['meses'] = MONTH_CHOICES
                    data['anio'] = AnioEjercicio.objects.filter(status=True).order_by('-anioejercicio')

                    data['estadoid'] = id_estado
                    data['idgrupo'] = idgrupo

                    return render(request, "helpdesk_hdplanificacion/viewcronograma.html", data)

                except Exception as ex:
                    pass

            if action == 'addplan':
                try:
                    data['title'] = u'Adicionar  Planificación'
                    form = HdPlanAprobacionForm()
                    data['perfilprin'] = perfilprin = request.session['perfilprincipal']
                    data['form'] = form
                    return render(request, "helpdesk_hdplanificacion/addplan.html", data)
                except Exception as ex:
                    pass
            if action == 'editplan':
                try:
                    data['title'] = u'Editar   Planificación'
                    data['plan'] = plan = HdPlanAprobacion.objects.get(pk=int(request.GET['id']))
                    data['perfilprin'] = perfilprin = request.session['perfilprincipal']
                    form = HdPlanAprobacionForm(initial={'fecharegistro': plan.fecharegistro,
                                                         'estadoaprobacion': plan.estadoaprobacion,
                                                         'solicitarevision': plan.solicitarevision,
                                                         })

                    if plan.periodo:
                        form.editar(plan.periodo)
                    data['form'] = form
                    return render(request, "helpdesk_hdplanificacion/editplan.html", data)
                except Exception as ex:
                    pass
            if action == 'delplan':
                try:
                    data['title'] = u'Eliminar   Planificación'
                    data['plan'] = HdPlanAprobacion.objects.get(pk=int(request.GET['id']))
                    return render(request, "helpdesk_hdplanificacion/deleteplan.html", data)
                except Exception as ex:
                    pass
            if action == 'viewplan':
                data['title'] = u'Registro  Planificación'
                search = None
                ids = None
                if 's' in request.GET:
                    search = request.GET['s']
                    ss = search.split(' ')
                    if len(ss) == 1:
                        plan = HdPlanAprobacion.objects.filter(periodo__anioejercicio__icontains=search, fecharegistro__icontains=search,
                                                               status=True,observacion__icontains=search)
                    elif len(ss) == 2:
                        plan = HdPlanAprobacion.objects.filter(Q(periodo__anioejercicio__icontains=ss[0]),
                                                               Q(observacion__icontains=ss[0]),Q(fecharegistro__icontains=ss[0]),
                                                               status=True)

                    else:
                        plan = HdPlanAprobacion.objects.filter(Q(periodo__anioejercicio__icontains=ss[0]),
                                                               Q(observacion__icontains=ss[0]),Q(fecharegistro__icontains=ss[0]),Q(periodo__anioejercicio__icontains=ss[1]),
                                                               Q(observacion__icontains=ss[1]),Q(fecharegistro__icontains=ss[1]),
                                                               Q(status=True))


                elif 'id' in request.GET:
                    ids = request.GET['id']
                    plan = HdPlanAprobacion.objects.filter(id=ids, status=True)
                else:
                    plan = HdPlanAprobacion.objects.filter(status=True)
                paging = MiPaginador(plan, 10)
                p = 1
                try:
                    paginasesion = 1
                    if 'paginador' in request.session:
                        paginasesion = int(request.session['paginador'])
                    if 'page' in request.GET:
                        p = int(request.GET['page'])
                    else:
                        p = paginasesion
                    try:
                        page = paging.page(p)
                    except:
                        p = 1
                    page = paging.page(p)
                except:
                    page = paging.page(p)
                request.session['paginador'] = p
                data['paging'] = paging
                data['rangospaging'] = paging.rangos_paginado(p)
                data['page'] = page
                data['search'] = search if search else ""
                data['ids'] = ids if ids else ""
                data['plan'] = page.object_list
                data['administrativo'] = persona
                return render(request, "helpdesk_hdplanificacion/viewplan.html", data)

            if action == 'buscargrupo':
                try:
                    q = request.GET['q'].upper().strip()
                    s = q.split(" ")


                    if s.__len__() == 1:
                        # activo = ActivoFijo.objects.filter(Q(codigogobierno__icontains=q) | Q(codigointerno__icontains=s[0]) | Q(serie__icontains=q),status=True, catalogo__equipoelectronico=True if tipo == 2 else False).distinct()[:20]
                        grupo = HdGrupoSistemaEquipo.objects.filter(
                            Q(descripcion__icontains=s[0]) , status=True).distinct()[:20]
                    else:
                        grupo = HdGrupoSistemaEquipo.objects.filter(
                            Q(descripcion__icontains=s[1]) ).filter(
                            status=True).distinct()[:20]



                    data = {"result": "ok","results": [{"id": x.id, "name": x.flexbox_reprhd()} for x in grupo]}
                    return JsonResponse(data)
                except Exception as ex:
                    pass

            elif action == 'buscarproveedor':
                try:
                    q = request.GET['q'].upper().strip()
                    s = q.split(" ")


                    if s.__len__() == 1:
                        # activo = ActivoFijo.objects.filter(Q(codigogobierno__icontains=q) | Q(codigointerno__icontains=s[0]) | Q(serie__icontains=q),status=True, catalogo__equipoelectronico=True if tipo == 2 else False).distinct()[:20]
                        proveedores = Proveedor.objects.filter((
                            Q(nombre__icontains=s[0]) | Q(identificacion__icontains=s[0])), status=True).distinct()[:20]
                    else:
                        proveedores = Proveedor.objects.filter((
                            Q(nombre__icontains=q)) & Q(status=True)).distinct()[:20]


                    data = {"result": "ok","results": [{"id": x.id, "identificacion": x.identificacion, "name": x.nombre} for x in proveedores]}
                    return JsonResponse(data)
                except Exception as ex:
                    pass

            elif action == 'activos_por_proveedor':
                try:
                    q = request.GET['q'].upper().strip()
                    s = q.split(" ")
                    filtrocatalogo = Q(status=True)
                    filtrocatalogo = filtrocatalogo & (Q(catalogo__equipoelectronico=True) &
                                                       Q(catalogo__status=True) &
                                                       Q(statusactivo=1) &
                                                       Q(status=True))
                    listadoingreso = IngresoProducto.objects.filter(proveedor_id=int(request.GET['idproveedor']),status=True).values_list('tipodocumento_id')
                    filtrocatalogo = filtrocatalogo & (Q(tipocomprobante_id__in=listadoingreso))
                    if s.__len__() == 1:
                        # activo = ActivoFijo.objects.filter(Q(codigogobierno__icontains=q) | Q(codigointerno__icontains=s[0]) | Q(serie__icontains=q),status=True, catalogo__equipoelectronico=True if tipo == 2 else False).distinct()[:20]
                        filtrocatalogo = filtrocatalogo & (Q(descripcion__icontains=s[0]) | Q(Q(codigogobierno__icontains=s[0])))
                    else:
                        filtrocatalogo = filtrocatalogo & (Q(descripcion__icontains=q) | Q(Q(codigogobierno__icontains=q)))

                    listado_activos = ActivoFijo.objects.filter(filtrocatalogo).distinct()[:20]

                    data = {"result": "ok","results": [{"id": x.id, "codigo": x.codigogobierno if x.codigogobierno else x.codigointerno, "name": x.catalogo.descripcion} for x in listado_activos]}
                    return JsonResponse(data)
                except Exception as ex:
                    pass


            elif action == 'buscarbien':
                try:
                    q = request.GET['q'].upper().strip()
                    s = q.split(" ")


                    if s.__len__() == 1:
                        # activo = ActivoFijo.objects.filter(Q(codigogobierno__icontains=q) | Q(codigointerno__icontains=s[0]) | Q(serie__icontains=q),status=True, catalogo__equipoelectronico=True if tipo == 2 else False).distinct()[:20]
                        bien = HdBien.objects.filter(
                            Q(sistemaequipo__icontains=s[0])|Q(cantidad__icontains=s[0]) |Q(observacion__icontains=s[0])|Q(gruposistema__descripcion__icontains=s[0]),status=True).distinct()[:20]
                    else:
                        bien = HdBien.objects.filter(
                            Q(sistemaequipo__icontains=s[1])| Q(cantidad__icontains=s[1])|
                            Q(observacion__icontains=s[1])| Q(gruposistema__descripcion__icontains=s[1]) ).filter(
                            status=True).distinct()[:20]



                    data = {"result": "ok","results": [{"id": x.id, "name": x.flexbox_reprhd()} for x in bien]}
                    return JsonResponse(data)
                except Exception as ex:
                    pass

            if action == 'buscarbienes':
                try:
                    q = request.GET['q'].upper().strip()
                    s = q.split(" ")
                    grupo = int(request.GET['grupo'])
                    activo = ActivoTecnologico.objects.filter(status=True)
                    if s.__len__() == 1:
                        # activo = ActivoFijo.objects.filter(Q(codigogobierno__icontains=q) | Q(codigointerno__icontains=s[0]) | Q(serie__icontains=q),status=True, catalogo__equipoelectronico=True if tipo == 2 else False).distinct()[:20]
                        bien = activo.filter(
                            Q(codigotic__icontains=s[0])|Q(codigogobierno__icontains=s[0]) |Q(codigointerno__icontains=s[0])|Q(observacion__icontains=s[0]),status=True,gruposistema=int(grupo)).distinct()[:20]
                    else:
                        bien = HdBien.objects.filter(
                            Q(sistemaequipo__icontains=s[1])| Q(cantidad__icontains=s[1])|
                            Q(observacion__icontains=s[1])| Q(gruposistema__descripcion__icontains=s[1]) ).filter(
                            status=True,gruposistema=int(grupo)).distinct()[:20]



                    data = {"result": "ok","results": [{"id": x.id, "name": x.flexbox_reprhd()} for x in bien]}
                    return JsonResponse(data)
                except Exception as ex:
                    pass

            if action == 'buscarfrecuencia':
                try:
                    q = request.GET['q'].upper().strip()
                    s = q.split(" ")


                    if s.__len__() == 1:
                        # activo = ActivoFijo.objects.filter(Q(codigogobierno__icontains=q) | Q(codigointerno__icontains=s[0]) | Q(serie__icontains=q),status=True, catalogo__equipoelectronico=True if tipo == 2 else False).distinct()[:20]
                        bien = HdConfFrecuencia.objects.filter(
                            Q(duracion__icontains=s[0])|Q(cantidad__icontains=s[0]) ,status=True).distinct()[:20]
                    else:
                        bien = HdConfFrecuencia.objects.filter(
                            Q(duracio__icontains=s[1])| Q(cantidad__icontains=s[1])
                             ).filter(
                            status=True).distinct()[:20]



                    data = {"result": "ok","results": [{"id": x.id, "name": x.flexbox_reprhd()} for x in bien]}
                    return JsonResponse(data)
                except Exception as ex:
                    pass
            if action == 'addcronograma':
                try:
                    data['title'] = u'Adicionar Cronograma Semanal '
                    form = HdCronogramaMantenimientoSemForm()
                    # data['repa'] = crono = HdMaterialMantenimiento.objects.get(pk=int(request.GET['reparacion']))
                    # data['materialesincidentes'] = incidente.hdmaterial_incidente_set.filter(status=True)

                    data['form'] = form
                    return render(request, "helpdesk_hdplanificacion/addcronograma.html", data)
                except Exception as ex:
                    pass
            if action == 'addactividad':
                try:
                    data['title'] = u'Agregar Actividades Semanales'
                    data['crosemana'] = mantsemanal = HdCronogramaMantenimientoSem.objects.get(pk=int(request.GET['id']))
                    grupo = None
                    form = HdCronogramaMantenimientoSemForm(initial={
                        'mes': mantsemanal.mes,

                    })
                    form.resolver()
                    if mantsemanal.gruposistema:
                        form.editar(mantsemanal.gruposistema)

                    data['crosemanales'] = HdDetCronogramaMantenimientoSem.objects.filter(
                        cronograma=int(mantsemanal.pk), status=True)
                    data['crosemanal'] = HdDetCronogramaMantenimientoSemForm()
                    if mantsemanal.gruposistema:
                        form.editar(mantsemanal.gruposistema)
                    data['form'] = form
                    return render(request, "helpdesk_hdplanificacion/editcronograma.html", data)
                except Exception as ex:
                    pass

            if action == 'bienpdf':
                try:
                    #tecnico = Administrativo.objects.get(pk=int(request.GET['tecnico'])).persona.pk
                    desde = request.GET['de']
                    hasta = request.GET['hasta']
                    grupo = int(request.GET['grupo'])

                    if grupo == 0:
                        bien = HdBien.objects.filter(Q(fecha_creacion__year=convertir_fecha(request.GET['de']).year)&
                                                     Q(fecha_creacion__month=convertir_fecha(request.GET['de']).month)&
                                                     Q(fecha_creacion__day=convertir_fecha(request.GET['de']).day)&
                                                     Q(fecha_creacion__year=convertir_fecha(request.GET['hasta']).year) &
                                                     Q(fecha_creacion__month=convertir_fecha(request.GET['hasta']).month) &
                                                     Q(fecha_creacion__day=convertir_fecha(request.GET['hasta']).day),
                                                     status=True)
                    else:
                        bien = HdBien.objects.filter(Q(fecha_creacion__year=convertir_fecha(request.GET['de']).year)&
                                                     Q(fecha_creacion__month=convertir_fecha(request.GET['de']).month)&
                                                     Q(fecha_creacion__day=convertir_fecha(request.GET['de']).day)&
                                                     Q(fecha_creacion__year=convertir_fecha(request.GET['hasta']).year) &
                                                     Q(fecha_creacion__month=convertir_fecha(request.GET['hasta']).month) &
                                                     Q(fecha_creacion__day=convertir_fecha(request.GET['hasta']).day),status=True, gruposistema=int(grupo))

                    if not grupo == 0:
                        data['grupo'] = HdGrupoSistemaEquipo.objects.get(pk=int(grupo)).descripcion
                    else:
                        data['grupo'] = 'TODOS LOS GRUPOS'


                    data['bien'] = bien
                    data['total'] = bien.count()
                    data['desde'] = desde
                    data['hasta'] = hasta

                    data['fechahoy'] = datetime.now().date()
                    return conviert_html_to_pdf(
                        'helpdesk_hdplanificacion/bienpdf.html',
                        {
                            'pagesize': 'A4',
                            'data': data,
                        }
                    )
                except Exception as ex:
                    pass

            if action == 'reportegeneral_bien':
                try:
                    tecnico=None
                    fechadesde = request.GET['de']
                    fechahasta = request.GET['hasta']
                    grupo = int(request.GET['grupo'])


                    if not grupo == 0:
                        data['grupo'] = HdBien.objects.get(gruposistema=int(grupo)).gruposistema.descripcion
                    else:
                        data['grupo'] = 'TODOS LOS GRUPOS'

                    __author__ = 'Unemi'
                    borders = Borders()
                    borders.left = 1
                    borders.right = 1
                    borders.top = 1
                    borders.bottom = 1
                    styrow = easyxf( 'font: name Times New Roman, color-index black, bold off; alignment: horiz centre; borders: left thin, right thin, top thin, bottom thin')
                    style_col = easyxf( 'pattern: pattern solid, fore_colour light_turquoise; font: name Times New Roman, color-index black, bold off; alignment: horiz centre; borders: left thin, right thin, top thin, bottom thin')
                    style_sb1 = easyxf('font: name Times New Roman, color-index black, bold on')
                    style_sb = easyxf('font: name Times New Roman, color-index black, bold off')
                    title = easyxf( 'font: name Times New Roman, color-index green, bold on , height 350; alignment: horiz centre')
                    title1 = easyxf( 'font: name Times New Roman, color-index green, bold on , height 250; alignment: horiz centre')
                    # style_date = easyxf('borders: left thin, right thin, top thin, bottom thin; alignment: horiz centre',
                    style_date = easyxf('borders: left thin, right thin, top thin, bottom thin; alignment: horiz centre',num_format_str='yy/mm/dd')
                    style1 = easyxf('borders: left thin, right thin, top thin, bottom thin; alignment: horiz centre',num_format_str='DD-MM-YYYY')
                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    # styrow.borders = borders
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('exp_xls_post_part')
                    # ti1=('TÉCNICO RESPONSABLE: ',style_sb)

                    ws.write_merge(0, 0, 0, 5, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                    ws.write_merge(1, 1, 0, 5, 'ÁREA DE MANTENIMIENTO', title1)
                    ws.write_merge(2, 2, 0, 5, 'REPORTE DE BIENES', title1)
                    ws.write_merge(4, 4, 0, 0, 'PERIODO:  ', style_sb1)
                    ws.write_merge(4, 4, 1, 1, 'DESDE  ' + request.GET['de'] + ' HASTA  ' + request.GET['hasta'],
                                   style_sb)
                    ws.write_merge(5, 5, 0, 0, 'GRUPO: ', style_sb1)
                    ws.write_merge(5, 5, 1, 1, data['grupo'], style_sb)
                    # ws.write_merge(6, 6, 0, 0, 'BLOQUE: ', style_sb1)
                    # ws.write_merge(6, 6, 1, 1, data['bloque'], style_sb)
                    # ws.write_merge(7, 7, 0, 0, 'ESTADO: ', style_sb1)
                    # ws.write_merge(7, 7, 1, 1, data['estado'], style_sb)
                    # ws.write_merge(3, 3, 0, 2, 'De ' +request.GET['de']+' hasta '+ request.GET['hasta'], style_sb)
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=Bienes ' + random.randint(1,10000).__str__() + '.xls'

                    columns = [
                        (u"N°", 4500),
                        (u"FECHA REGISTRO", 9000),
                        (u"UBICACIÓN", 10000),
                        (u"GRUPO", 10000),
                        (u"CANTIDAD", 10000),
                        (u"OBSERVACIÓN", 11000),

                    ]
                    row_num = 9
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], style_col)
                        ws.col(col_num).width = columns[col_num][1]
                    date_format = xlwt.XFStyle()
                    date_format.num_format_str = 'yyyy/mm/dd'

                    if grupo == 0:
                        bien = HdBien.objects.filter(Q(fecha_creacion__year=convertir_fecha(request.GET['de']).year) &
                                                     Q(fecha_creacion__month=convertir_fecha(request.GET['de']).month) &
                                                     Q(fecha_creacion__day=convertir_fecha(request.GET['de']).day) &
                                                     Q(fecha_creacion__year=convertir_fecha(
                                                         request.GET['hasta']).year) &
                                                     Q(fecha_creacion__month=convertir_fecha(
                                                         request.GET['hasta']).month) &
                                                     Q(fecha_creacion__day=convertir_fecha(request.GET['hasta']).day),
                                                     status=True)
                    else:
                        bien = HdBien.objects.filter(Q(fecha_creacion__year=convertir_fecha(request.GET['de']).year) &
                                                     Q(fecha_creacion__month=convertir_fecha(request.GET['de']).month) &
                                                     Q(fecha_creacion__day=convertir_fecha(request.GET['de']).day) &
                                                     Q(fecha_creacion__year=convertir_fecha(
                                                         request.GET['hasta']).year) &
                                                     Q(fecha_creacion__month=convertir_fecha(
                                                         request.GET['hasta']).month) &
                                                     Q(fecha_creacion__day=convertir_fecha(request.GET['hasta']).day),
                                                     status=True, gruposistema=int(grupo))
                    row_num = 10
                    i=0
                    for bienes in bien:
                        i+=1
                        ws.write(row_num, 0,i, styrow)
                        ws.write(row_num, 1, (bienes.fecha_creacion), style1)
                        ws.write(row_num, 2, str(bienes.ubicacion.bloque.nombre)+'-'+str(bienes.ubicacion.ubicacion.nombre), styrow)
                        ws.write(row_num, 3, str(bienes.gruposistema.descripcion), styrow)
                        ws.write(row_num, 4, str(bienes.cantidad), styrow)
                        ws.write(row_num, 5, str(bienes.observacion), styrow)
                        row_num += 1
                    row_num += 1
                    ws.write(row_num, 4, 'TOTAL', styrow)
                    ws.write(row_num, 5, bien.count(), styrow)
                    wb.save(response)
                    return response
                except Exception as ex:
                    pass

            if action == 'frecuenciapdf':
                try:

                    grupo = int(request.GET['grupo'])

                    if grupo == 0:
                        frecuencia = HdFrecuencia.objects.filter( status=True)
                    else:
                        frecuencia = HdFrecuencia.objects.filter(status=True,gruposistema=int(grupo))

                    if not grupo == 0:
                        data['grupo'] = HdFrecuencia.objects.get(pk=int(grupo)).gruposistema.descripcion
                    else:
                        data['grupo'] = 'TODOS LOS GRUPOS'

                    data['frecuencia'] = frecuencia
                    data['total'] = frecuencia.count()
                    data['fechahoy'] = datetime.now().date()
                    return conviert_html_to_pdf(
                        'helpdesk_hdplanificacion/frecuenciapdf.html',
                        {
                            'pagesize': 'A4 landscape',
                            'data': data,

                        }
                    )
                except Exception as ex:
                    pass

            if action == 'reportegeneral_frecuencia':
                try:


                    grupo = int(request.GET['grupo'])

                    if not grupo == 0:
                        data['grupo'] = HdFrecuencia.objects.get(gruposistema=int(grupo)).gruposistema.descripcion
                    else:
                        data['grupo'] = 'TODOS LOS GRUPOS'

                    __author__ = 'Unemi'
                    borders = Borders()
                    borders.left = 1
                    borders.right = 1
                    borders.top = 1
                    borders.bottom = 1
                    styrow = easyxf(
                        'font: name Times New Roman, color-index black, bold off; alignment: horiz centre; borders: left thin, right thin, top thin, bottom thin')
                    style_col = easyxf(
                        'pattern: pattern solid, fore_colour light_turquoise; font: name Times New Roman, color-index black, bold off; alignment: horiz centre; borders: left thin, right thin, top thin, bottom thin')
                    style_sb1 = easyxf('font: name Times New Roman, color-index black, bold on')
                    style_sb = easyxf('font: name Times New Roman, color-index black, bold off')
                    title = easyxf(
                        'font: name Times New Roman, color-index green, bold on , height 350; alignment: horiz centre')
                    title1 = easyxf(
                        'font: name Times New Roman, color-index green, bold on , height 250; alignment: horiz centre')
                    # style_date = easyxf('borders: left thin, right thin, top thin, bottom thin; alignment: horiz centre',
                    style_date = easyxf(
                        'borders: left thin, right thin, top thin, bottom thin; alignment: horiz centre',
                        num_format_str='yy/mm/dd')
                    style1 = easyxf('borders: left thin, right thin, top thin, bottom thin; alignment: horiz centre',
                                    num_format_str='DD-MM-YYYY')
                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    # styrow.borders = borders
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('exp_xls_post_part')
                    # ti1=('TÉCNICO RESPONSABLE: ',style_sb)

                    ws.write_merge(0, 0, 0, 7, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                    ws.write_merge(1, 1, 0, 7, 'ÁREA DE MANTENIMIENTO', title1)
                    ws.write_merge(2, 2, 0, 7, 'FRECUENCIA DE MANTENIMIENTOS', title1)
                    # ws.write_merge(4, 4, 0, 0, 'PERIODO:  ', style_sb1)
                    # ws.write_merge(4, 4, 1, 1, 'DESDE  ' + request.GET['de'] + ' HASTA  ' + request.GET['hasta'],
                    #                style_sb)
                    ws.write_merge(4, 4, 0, 0, 'GRUPO: ', style_sb1)
                    ws.write_merge(4, 4, 1, 1, data['grupo'], style_sb)
                    # ws.write_merge(6, 6, 0, 0, 'BLOQUE: ', style_sb1)
                    # ws.write_merge(6, 6, 1, 1, data['bloque'], style_sb)
                    # ws.write_merge(7, 7, 0, 0, 'ESTADO: ', style_sb1)
                    # ws.write_merge(7, 7, 1, 1, data['estado'], style_sb)
                    # ws.write_merge(3, 3, 0, 2, 'De ' +request.GET['de']+' hasta '+ request.GET['hasta'], style_sb)
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=FrecuenciaMantenimientos ' + random.randint(1,
                                                                                                      10000).__str__() + '.xls'

                    columns = [
                        (u"N°", 4500),
                        (u"GRUPO", 7000),
                        (u"SISTEMAS/EQUIPOS DE MANTENIMIENTO", 9000),
                        (u"TIPO DE MANTENIMIENTO", 10000),
                        (u"DESCRIPCIÓN DE TRABAJO", 10000),
                        (u"FRECUENCIA", 5000),
                        (u"CONSIDERACIONES ESP.", 11000),
                        (u"PROCESO", 11000),

                    ]
                    row_num = 6
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], style_col)
                        ws.col(col_num).width = columns[col_num][1]
                    date_format = xlwt.XFStyle()
                    date_format.num_format_str = 'yyyy/mm/dd'

                    if grupo == 0:
                        frecuencias = HdFrecuencia.objects.filter( status=True)
                    else:
                        frecuencias = HdFrecuencia.objects.filter(status=True,gruposistema=int(grupo))

                    row_num = 7
                    i = 0
                    for frecuencia in frecuencias:
                        i += 1
                        ws.write(row_num, 0, i, styrow)
                        ws.write(row_num, 1, (frecuencia.bien.gruposistema.descripcion), style1)
                        ws.write(row_num, 2, (frecuencia.bien.sistemaequipo), style1)
                        ws.write(row_num, 3,str(frecuencia.get_tipomantenimiento_display()),styrow)
                        ws.write(row_num, 4,str(frecuencia.descripcion),styrow)
                        ws.write(row_num, 5,str(frecuencia.frecuencia.cantidad)+' '+str(frecuencia.frecuencia.get_duracion_display()),styrow)
                        ws.write(row_num, 6,str(frecuencia.consideracion),styrow)
                        ws.write(row_num, 7,str(frecuencia.get_proceso_display()),styrow)
                        row_num += 1
                    row_num += 1
                    ws.write(row_num, 6, 'TOTAL', styrow)
                    ws.write(row_num, 7, frecuencias.count(), styrow)
                    wb.save(response)
                    return response
                except Exception as ex:
                    pass

            if action == 'cronogramapdf':
                try:
                    cronograma = HdCronogramaMantenimiento.objects.get(id=int(request.GET['id']))
                    data['mantenimientosrealizados'] = mantenimientosrealizados = HdDetMantenimientosActivos.objects.filter(status=True, cronograma=cronograma)


                    data['fechahoy'] = datetime.now().date()
                    return conviert_html_to_pdf(
                        'helpdesk_hdplanificacion/cronogramapdf.html',
                        {
                            'pagesize': 'A4',
                            'data': data,
                        }
                    )
                except Exception as ex:
                    pass

            if action == 'reportegeneral_cronograma':
                try:

                    __author__ = 'Unemi'
                    style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
                    style_nb = easyxf('font: name Times New Roman, color-index blue, bold on',
                                      num_format_str='#,##0.00')
                    style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
                    title = easyxf(
                        'font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                    style1 = easyxf(num_format_str='D-MMM-YY')
                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('exp_xls_post_part')
                    # ws.write_merge(0, 0, 0, 7, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=listado_mantenimientos' + random.randint(1,
                                                                                                              10000).__str__() + '.xls'
                    row_num = 0
                    columns = [
                        (u"N.", 2000),
                        (u"Responsable", 3000),
                        (u"Fecha", 6000),
                        (u"Sistemas/Equipos", 6000),
                        (u"Observacion", 6000),
                        (u"Cod. Gob.", 6000),
                        (u"Cod. Int.", 6000),
                        (u"Cod. Tics.", 6000),

                    ]
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], font_style)
                        ws.col(col_num).width = columns[col_num][1]
                    date_format = xlwt.XFStyle()
                    date_format.num_format_str = 'yyyy/mm/dd'
                    resultado = None
                    cronograma = HdCronogramaMantenimiento.objects.get(id=int(request.GET['id']))
                    data['mantenimientosrealizados'] = mantenimientosrealizados = HdDetMantenimientosActivos.objects.filter(status=True, cronograma=cronograma)
                    row_num = 0
                    for mantenimiento in mantenimientosrealizados:
                        row_num += 1
                        ws.write(row_num, 0, row_num, font_style2)
                        ws.write(row_num, 1, mantenimiento.activotecno.activotecnologico.responsable.__str__(), font_style2)
                        ws.write(row_num, 2, mantenimiento.fecha, font_style2)
                        ws.write(row_num, 3, mantenimiento.activotecno.__str__(), font_style2)
                        ws.write(row_num, 4, mantenimiento.observaciones, font_style2)
                        ws.write(row_num, 5, mantenimiento.activotecno.codigogobierno, font_style2)
                        ws.write(row_num, 6, mantenimiento.activotecno.codigointerno, font_style2)
                        ws.write(row_num, 7, mantenimiento.activotecno.codigotic, font_style2)

                        # ws.write(row_num, 7, campo7, font_style2)
                    wb.save(response)
                    return response
                except Exception as ex:
                    pass

            if action == 'reportegeneral_cronogramasem':
                try:

                    mes = int(request.GET['mes'])
                    anio = AnioEjercicio.objects.get(pk=request.GET['anio']).anioejercicio
                    grupo = int(request.GET['grupo'])

                    if not grupo == 0:
                        gr = ' '
                        gru = HdDetCronogramaMantenimientoSem.objects.filter(cronograma__gruposistema=int(grupo),
                                                                          fecha_creacion__year=anio,
                                                                          cronograma__mes=int(mes)).distinct('cronograma')
                        for x in gru:
                            gr = x.cronograma.gruposistema.descripcion
                        data['grupo'] = gr
                    else:
                        data['grupo'] = 'TODOS LOS GRUPOS'

                    if not mes == 0:
                        for x in MONTH_CHOICES:
                            if x[0] == int(mes):
                                data['mes'] = x[1]
                                print(x[1])
                    else:
                        data['mes'] = 'TODOS '

                    __author__ = 'Unemi'
                    borders = Borders()
                    borders.left = 1
                    borders.right = 1
                    borders.top = 1
                    borders.bottom = 1
                    styrow = easyxf(
                        'font: name Times New Roman, color-index black, bold off; alignment: horiz centre; borders: left thin, right thin, top thin, bottom thin')
                    styrow1 = easyxf(
                        'font: name Times New Roman, color-index black, bold off; alignment: horiz centre;')
                    style_col = easyxf(
                        'pattern: pattern solid, fore_colour light_turquoise; font: name Times New Roman, color-index black, bold off; alignment: horiz centre; borders: left thin, right thin, top thin, bottom thin')
                    style_sb1 = easyxf('font: name Times New Roman, color-index black, bold on')
                    style_sb = easyxf('font: name Times New Roman, color-index black, bold off')
                    title = easyxf(
                        'font: name Times New Roman, color-index green, bold on , height 350; alignment: horiz centre')
                    title1 = easyxf(
                        'font: name Times New Roman, color-index green, bold on , height 250; alignment: horiz centre')
                    # style_date = easyxf('borders: left thin, right thin, top thin, bottom thin; alignment: horiz centre',
                    style_date = easyxf(
                        'borders: left thin, right thin, top thin, bottom thin; alignment: horiz centre',
                        num_format_str='yy/mm/dd')
                    style1 = easyxf('borders: left thin, right thin, top thin, bottom thin; alignment: horiz centre',
                                    num_format_str='DD-MM-YYYY')
                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    # styrow.borders = borders
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('exp_xls_post_part')

                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=CronogramaMensual ' + random.randint(1,10000).__str__() + '.xls'

                    if grupo == 0 and mes == 0:
                        mensual = HdDetCronogramaMantenimientoSem.objects.filter(fecha_creacion__year=anio, status=True)
                    elif not grupo == 0 and not mes == 0:
                        mensual = HdDetCronogramaMantenimientoSem.objects.filter(cronograma__mes=int(mes), fecha_creacion__year=anio,
                                                                              status=True,
                                                                              cronograma__gruposistema=int(grupo))
                    elif not grupo == 0 and mes == 0:
                        mensual = HdDetCronogramaMantenimientoSem.objects.filter(cronograma__gruposistema=int(grupo),
                                                                              fecha_creacion__year=anio,
                                                                              status=True)
                    elif grupo == 0 and not mes == 0:
                        mensual = HdDetCronogramaMantenimientoSem.objects.filter(fecha_creacion__year=anio, cronograma__mes=int(mes),
                                                                              status=True,
                                                                              )
                    bien =mensual.distinct('fechainicio')
                    columns = [
                        (u"UBICACIÓN", 15000),
                    ]
                    for x in bien:

                        columns.append([str(x.fechainicio.day),str(x.fechafin.day)])
                        columns.append(['CANTIDAD',7000])

                    row_num = 4
                    for col_num in range(len(columns)):
                        if col_num > 0  and not str(columns[col_num][0])=='CANTIDAD':
                            ws.write(row_num, col_num, 'SEMANA '+str(columns[col_num][0])+' - '+str(columns[col_num][1]), style_col)
                            ws.col(col_num).width = 8000

                        elif col_num==0 or str(columns[col_num][0])=='CANTIDAD':
                            ws.write(row_num, col_num, columns[col_num][0], style_col)
                            ws.col(col_num).width = columns[col_num][1]


                    date_format = xlwt.XFStyle()
                    date_format.num_format_str = 'yyyy/mm/dd'
                    ini = 5
                    row_num = 5
                    i = 0
                    y = 5
                    p = 0
                    bloque = mensual.distinct('bloque__bloque')
                    for x in bloque:
                        c = 0
                        for mes in mensual:
                                for p in range(len(columns)):
                                    if p>0:
                                        if str(columns[p][0]) == str(mes.fechainicio.day) and str(columns[p][1])==str(mes.fechafin.day) and x.bloque.bloque.pk==mes.bloque.bloque.pk:
                                            ws.write(y, p, str(mes.bien.sistemaequipo), styrow)
                                            p += 1
                                            ws.write(y, p, int(mes.cantidad), styrow)
                                            y += 1
                                            row_num += 1
                                            c += 1
                        i+=1
                        if i == 1:
                            e = y - 1
                            ws.write_merge(ini, e, 0, 0, (x.bloque.bloque.nombre), styrow1)
                        else:
                            d = y - (c)
                            f = y - 1
                            ws.write_merge(d, f, 0, 0, (x.bloque.bloque.nombre), styrow1)
                    ws.write_merge(0, 0, 0, p, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                    ws.write_merge(1, 1, 0, p, 'ÁREA DE MANTENIMIENTO', title1)
                    ws.write_merge(2, 2, 0, p, 'CRONOGRAMA DE  MANTENIMIENTO' + ' ' + data['grupo'], title1)
                    # i += 1
                    # ws.write(row_num, 0, i, styrow)
                    # ws.write(row_num, 1, (mes.get_mes_display()), styrow)
                    #
                    # ws.write(row_num, 3, str(mes.bien.sistemaequipo), styrow)
                    # ws.write(row_num, 4, str(mes.cantidad), styrow)


                    row_num += 1
                    # ws.write(row_num, 1, 'TOTAL', styrow1)
                    # alfabeto = [(0, 'A'), (1, 'B'), (2, 'C'), (3, 'D'), (4, 'E'), (5, 'F'), (6, 'G'), (7, 'H'),
                    #             (8, 'I'), (9, 'J'), (10, 'K'), (11, 'L'), (12, 'M'),
                    #             (13, 'N'), (14, 'O'), (15, 'P'), (16, 'Q'), (17, 'R'), (18, 'S'), (19, 'T'), (20, 'U'),
                    #             (21, 'V'), (22, 'W'), (23, 'X'), (24, 'Y'), (25, 'Z')]
                    #
                    # for p in range(len(columns)):
                    #     if p > 1:
                    #         for x in range(len(alfabeto)):
                    #
                    #             if alfabeto[x][0] == p:
                    #                 ws.write(row_num, p, Formula(
                    #                     "SUM(" + str(alfabeto[x][1]) + "11:" + str(alfabeto[x][1]) + str(
                    #                         row_num) + ")"), styrow1)

                    wb.save(response)
                    return response
                except Exception as ex:
                    pass
            if action == 'cronogramasempdf':
                try:
                    mes = int(request.GET['mes'])
                    anio = AnioEjercicio.objects.get(pk=request.GET['anio']).anioejercicio
                    grupo = int(request.GET['grupo'])

                    if not grupo == 0:
                        gr = ' '
                        gru = HdDetCronogramaMantenimientoSem.objects.filter(cronograma__gruposistema=int(grupo),
                                                                             fecha_creacion__year=anio,
                                                                             cronograma__mes=int(mes)).distinct(
                            'cronograma')
                        for x in gru:
                            gr = x.cronograma.gruposistema.descripcion
                        data['grupo'] = gr
                    else:
                        data['grupo'] = 'TODOS LOS GRUPOS'



                    if not mes == 0:
                        for x in MONTH_CHOICES:
                            if x[0] == int(mes):
                                data['meses'] = x[1]
                                print(x[1])
                    else:
                        data['meses'] = 'TODOS '
                    data['anio'] = anio
                    mensual = None
                    if grupo == 0 and mes == 0:
                        mensual = HdDetCronogramaMantenimientoSem.objects.filter(fecha_creacion__year=anio, status=True)
                    elif not grupo == 0 and not mes == 0:
                        mensual = HdDetCronogramaMantenimientoSem.objects.filter(cronograma__mes=int(mes),
                                                                                 fecha_creacion__year=anio,
                                                                                 status=True,
                                                                                 cronograma__gruposistema=int(grupo))
                    elif not grupo == 0 and mes == 0:
                        mensual = HdDetCronogramaMantenimientoSem.objects.filter(cronograma__gruposistema=int(grupo),
                                                                                 fecha_creacion__year=anio,
                                                                                 status=True)
                    elif grupo == 0 and not mes == 0:
                        mensual = HdDetCronogramaMantenimientoSem.objects.filter(fecha_creacion__year=anio,
                                                                                 cronograma__mes=int(mes),
                                                                                 status=True,
                                                                                 )
                    data['bien'] = bien = mensual.distinct('fechainicio')
                    data['bloque'] =bloque = mensual.distinct('bloque__bloque')

                    data['mensual'] = mensual


                    data['fechahoy'] = datetime.now().date()
                    return conviert_html_to_pdf(
                        'helpdesk_hdplanificacion/cronogramasempdf.html',
                        {
                            'pagesize': 'A4',
                            'data': data,
                        }
                    )
                except Exception as ex:
                    pass
            if action == 'reportegeneral_presupuesto':
                try:
                    anio = AnioEjercicio.objects.get(pk=request.GET['anio']).anioejercicio
                    grupo = int(request.GET['grupo'])
                    if not grupo == 0:
                        gr = ' '
                        gru = HdPresupuestoRecurso.objects.filter(gruposistema=int(grupo),fecha_creacion__year=anio)
                        for x in gru:
                            gr = x.gruposistema.descripcion
                        data['grupo'] = gr
                    else:
                        data['grupo'] = 'TODOS LOS GRUPOS'
                    __author__ = 'Unemi'
                    borders = Borders()
                    borders.left = 1
                    borders.right = 1
                    borders.top = 1
                    borders.bottom = 1
                    styrow = easyxf('font: name Times New Roman, color-index black, bold off; alignment: horiz centre; borders: left thin, right thin, top thin, bottom thin')
                    styrow1 = easyxf('font: name Times New Roman, color-index black, bold off; alignment: horiz centre;')
                    style_col = easyxf('pattern: pattern solid, fore_colour light_turquoise; font: name Times New Roman, color-index black, bold off; alignment: horiz centre; borders: left thin, right thin, top thin, bottom thin')
                    style_sb1 = easyxf('font: name Times New Roman, color-index black, bold on')
                    style_sb = easyxf('font: name Times New Roman, color-index black, bold off')
                    title = easyxf('font: name Times New Roman, color-index green, bold on , height 350; alignment: horiz centre')
                    title1 = easyxf('font: name Times New Roman, color-index green, bold on , height 250; alignment: horiz centre')
                    style_date = easyxf('borders: left thin, right thin, top thin, bottom thin; alignment: horiz centre',num_format_str='yy/mm/dd')
                    style1 = easyxf('borders: left thin, right thin, top thin, bottom thin; alignment: horiz centre',num_format_str='DD-MM-YYYY')
                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('exp_xls_post_part')
                    ws.write_merge(0, 0, 0, 4, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                    ws.write_merge(1, 1, 0, 4, 'ÁREA DE MANTENIMIENTO', title1)
                    ws.write_merge(2, 2, 0, 4, 'RECURSO PRESUPUESTOS' + ' ' + data['grupo'], title1)
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=RecursoPresupuesto ' + random.randint(1,10000).__str__() + '.xls'
                    if grupo == 0 :
                        mensual = HdPresupuestoRecurso.objects.filter(fecha_creacion__year=anio, status=True)
                    else:
                        mensual = HdPresupuestoRecurso.objects.filter(fecha_creacion__year=anio, status=True,cronograma__gruposistema=int(grupo))
                    columns = [
                        (u"GRUPO", 4500),
                        (u"TIPO", 10000),
                        (u"PRESUPUESTO PROYECTADO REQUERIDO", 10000),
                        (u"PRESUPUESTO PROYECTADO INCL. IVA", 10000),
                    ]
                    row_num = 5
                    for col_num in range(len(columns)):
                            ws.write(row_num, col_num, columns[col_num][0], style_col)
                            ws.col(col_num).width = columns[col_num][1]
                    date_format = xlwt.XFStyle()
                    date_format.num_format_str = 'yyyy/mm/dd'
                    row_num = 6
                    for x in mensual:
                        ws.write(row_num, 0, str(x.gruposistema.descripcion), styrow1)
                        ws.write(row_num, 1, str(x.bien.sistemaequipo), styrow1)
                        ws.write(row_num, 2, str(x.presupuestoreq), styrow1)
                        ws.write(row_num, 3, str(x.presupuestototal), styrow1)
                        row_num += 1
                    wb.save(response)
                    return response
                except Exception as ex:
                    pass
            if action == 'reportegeneral_reparacion':
                try:
                    # anio = AnioEjercicio.objects.get(pk=request.GET['anio']).anioejercicio
                    grupo = int(request.GET['grupo'])
                    bloque = int(request.GET['bloque'])
                    ubicacion = int(request.GET['ubicacion'])
                    if not grupo == 0:
                        gr = ' '
                        gru = HdReparacion.objects.filter(gruposistema=int(grupo),ubicacion__bloque__pk=int(bloque),ubicacion__ubicacion__pk=int(ubicacion),usuario_creacion=int(persona.usuario.id))
                        for x in gru:
                            gr = x.gruposistema.descripcion
                        data['grupo'] = gr
                    else:
                        data['grupo'] = 'TODOS LOS GRUPOS'
                    __author__ = 'Unemi'
                    borders = Borders()
                    borders.left = 1
                    borders.right = 1
                    borders.top = 1
                    borders.bottom = 1
                    styrow = easyxf('font: name Times New Roman, color-index black, bold off; alignment: horiz centre; borders: left thin, right thin, top thin, bottom thin')
                    styrow1 = easyxf('font: name Times New Roman, color-index black, bold off; alignment: horiz centre;')
                    style_col = easyxf('pattern: pattern solid, fore_colour light_turquoise; font: name Times New Roman, color-index black, bold off; alignment: horiz centre; borders: left thin, right thin, top thin, bottom thin')
                    style_sb1 = easyxf('font: name Times New Roman, color-index black, bold on')
                    style_sb = easyxf('font: name Times New Roman, color-index black, bold off')
                    title = easyxf('font: name Times New Roman, color-index green, bold on , height 350; alignment: horiz centre')
                    title1 = easyxf('font: name Times New Roman, color-index green, bold on , height 250; alignment: horiz centre')
                    style_date = easyxf('borders: left thin, right thin, top thin, bottom thin; alignment: horiz centre',num_format_str='yy/mm/dd')
                    style1 = easyxf('borders: left thin, right thin, top thin, bottom thin; alignment: horiz centre',num_format_str='DD-MM-YYYY')
                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('exp_xls_post_part')
                    ws.write_merge(0, 0, 0, 4, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                    ws.write_merge(1, 1, 0, 4, 'REGISTRO DE NECESIDADES MANTENIMIENTOS', title1)
                    ws.write_merge(2, 2, 0, 4, '' + ' ' + data['grupo'], title1)
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=RecursoPresupuesto ' + random.randint(1,10000).__str__() + '.xls'
                    if grupo == 0 :
                        mensual = HdDetalle_Reparacion.objects.filter(reparacion__ubicacion__bloque__pk=int(bloque),reparacion__ubicacion__ubicacion__pk=int(ubicacion),usuario_creacion=int(persona.usuario.id))
                    else:
                        mensual =HdDetalle_Reparacion.objects.filter(reparacion__gruposistema=int(grupo),reparacion__ubicacion__bloque__pk=int(bloque),reparacion__ubicacion__ubicacion__pk=int(ubicacion),usuario_creacion=int(persona.usuario.id))
                    columns = [
                        (u"SISTEMA /EQUIPOS", 4500),
                        (u"EXISTENCIA", 10000),
                        (u"DESCRIPCIÓN DEL TRABAJO", 10000),
                        (u"CANTIDAD EN LEVANTAMIENTO", 10000),
                    ]
                    row_num = 5
                    for col_num in range(len(columns)):
                            ws.write(row_num, col_num, columns[col_num][0], style_col)
                            ws.col(col_num).width = columns[col_num][1]
                    date_format = xlwt.XFStyle()
                    date_format.num_format_str = 'yyyy/mm/dd'
                    row_num = 6
                    for x in mensual:
                        ws.write(row_num, 0, str(x.bien.sistemaequipo), styrow1)
                        ws.write(row_num, 1, str(x.bien.cantidad), styrow1)
                        ws.write(row_num, 2, str(x.descripcion), styrow1)
                        ws.write(row_num, 3, str(x.cantidad), styrow1)
                        row_num += 1
                    row_num += 3
                    tecnico=Persona.objects.get(pk=int(persona.id))
                    ws.write(row_num, 2, '________________________________________________', styrow1)
                    row_num += 1
                    ws.write(row_num, 2, str(tecnico.nombre_completo_inverso()), styrow1)
                    row_num += 1
                    ws.write(row_num, 2, str(tecnico.mi_cargo()), styrow1)

                    wb.save(response)
                    return response
                except Exception as ex:
                    pass
            if action == 'reportepdfreparacion':
                try:
                    grupo = int(request.GET['grupo'])
                    bloque = int(request.GET['bloque'])
                    ubicacion = int(request.GET['ubicacion'])
                    if not grupo == 0:
                        gr = ' '
                        gru = HdReparacion.objects.filter(gruposistema=int(grupo), ubicacion__bloque__pk=int(bloque),
                                                          ubicacion__ubicacion__pk=int(ubicacion),
                                                          usuario_creacion=int(persona.usuario.id))
                        for x in gru:
                            gr = x.gruposistema.descripcion
                        data['grupo'] = gr
                    else:
                        data['grupo'] = 'TODOS LOS GRUPOS'
                    if grupo == 0 :
                        mensual = HdDetalle_Reparacion.objects.filter(reparacion__ubicacion__bloque__pk=int(bloque),reparacion__ubicacion__ubicacion__pk=int(ubicacion),usuario_creacion=int(persona.usuario.id))
                    else:
                        mensual =HdDetalle_Reparacion.objects.filter(reparacion__gruposistema=int(grupo),reparacion__ubicacion__bloque__pk=int(bloque),reparacion__ubicacion__ubicacion__pk=int(ubicacion),usuario_creacion=int(persona.usuario.id))


                    data['reparacion'] = mensual
                    data['tecnico'] =  tecnico = Persona.objects.get(pk=int(persona.id))
                    data['bloque'] = HdBloqueUbicacion.objects.get(status=True,bloque=bloque,ubicacion=ubicacion)

                    data['fechahoy'] = datetime.now().date()
                    return conviert_html_to_pdf(
                        'helpdesk_hdplanificacion/reparacionpdf.html',
                        {
                            'pagesize': 'A4',
                            'data': data,
                        }
                    )
                except Exception as ex:
                    pass

            if action == 'reportegeneralconsolidado':
                try:
                    # anio = AnioEjercicio.objects.get(pk=request.GET['anio']).anioejercicio
                    grupo = int(request.GET['grupo'])
                    bloque = int(request.GET['bloque'])
                    ubicacion = int(request.GET['ubicacion'])
                    tecnico = int(request.GET['tecnico'])

                    if not grupo == 0:
                        gr = ' '
                        gru = HdReparacion.objects.filter(gruposistema=int(grupo),ubicacion__bloque__pk=int(bloque),ubicacion__ubicacion__pk=int(ubicacion),usuario_creacion=int(persona.usuario.id))
                        for x in gru:
                            gr = x.gruposistema.descripcion
                        data['grupo'] = gr
                    else:
                        data['grupo'] = 'TODOS LOS GRUPOS'
                    if not tecnico == 0:
                        tecnico = Persona.objects.get(status=True, pk=int(request.GET['tecnico'])).usuario.pk

                    else:
                        data['grupo'] = 'TODOS LOS TÉCNICOS'
                    __author__ = 'Unemi'
                    borders = Borders()
                    borders.left = 1
                    borders.right = 1
                    borders.top = 1
                    borders.bottom = 1
                    styrow = easyxf('font: name Times New Roman, color-index black, bold off; alignment: horiz centre; borders: left thin, right thin, top thin, bottom thin')
                    styrow1 = easyxf('font: name Times New Roman, color-index black, bold off; alignment: horiz centre;')
                    style_col = easyxf('pattern: pattern solid, fore_colour light_turquoise; font: name Times New Roman, color-index black, bold off; alignment: horiz centre; borders: left thin, right thin, top thin, bottom thin')
                    style_sb1 = easyxf('font: name Times New Roman, color-index black, bold on')
                    style_sb = easyxf('font: name Times New Roman, color-index black, bold off')
                    title = easyxf('font: name Times New Roman, color-index green, bold on , height 350; alignment: horiz centre')
                    title1 = easyxf('font: name Times New Roman, color-index green, bold on , height 250; alignment: horiz centre')
                    style_date = easyxf('borders: left thin, right thin, top thin, bottom thin; alignment: horiz centre',num_format_str='yy/mm/dd')
                    style1 = easyxf('borders: left thin, right thin, top thin, bottom thin; alignment: horiz centre',num_format_str='DD-MM-YYYY')
                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('exp_xls_post_part')
                    ws.write_merge(0, 0, 0, 5, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                    ws.write_merge(1, 1, 0, 5, 'CONSOLIDADO  DE REGISTRO DE NECESIDADES MANTENIMIENTOS', title1)
                    ws.write_merge(2, 2, 0, 5, '' + ' ' + data['grupo'], title1)
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=ConsolidadoNecesidades ' + random.randint(1,10000).__str__() + '.xls'
                    if grupo == 0 and tecnico ==0 and bloque == 0 and ubicacion ==0:
                        mensual = HdDetalle_Reparacion.objects.filter(status=True)
                    elif not grupo == 0 and  not tecnico ==0 and  not bloque == 0 and  not ubicacion ==0:
                        mensual =HdDetalle_Reparacion.objects.filter(reparacion__gruposistema=int(grupo),reparacion__ubicacion__bloque__pk=int(bloque),reparacion__ubicacion__ubicacion__pk=int(ubicacion),usuario_creacion=int(tecnico))
                    elif  not  grupo == 0 and  tecnico == 0 and bloque == 0 and ubicacion ==0:
                        mensual = HdDetalle_Reparacion.objects.filter(reparacion__gruposistema=int(grupo))
                    elif  not  grupo == 0 and  not tecnico == 0 and bloque == 0 and ubicacion ==0:
                        mensual = HdDetalle_Reparacion.objects.filter(reparacion__gruposistema=int(grupo),usuario_creacion=int(tecnico))
                    elif  not  grupo == 0 and not tecnico == 0 and  not bloque == 0 and ubicacion ==0:
                        mensual = HdDetalle_Reparacion.objects.filter(reparacion__gruposistema=int(grupo),reparacion__ubicacion__bloque__pk=int(bloque),usuario_creacion=int(tecnico))
                    elif    grupo == 0 and  not  tecnico == 0 and bloque == 0 and ubicacion ==0:
                        mensual = HdDetalle_Reparacion.objects.filter(usuario_creacion=int(tecnico))
                    elif  grupo == 0 and  tecnico == 0 and  not bloque == 0 and ubicacion ==0:
                        mensual = HdDetalle_Reparacion.objects.filter( reparacion__ubicacion__bloque__pk=int(bloque) )
                    elif  grupo == 0 and  tecnico == 0 and   bloque == 0 and not ubicacion ==0:
                        mensual = HdDetalle_Reparacion.objects.filter( reparacion__ubicacion__ubicacion__pk=int(ubicacion) )

                    columns = [
                        (u"USUARIO", 4500),
                        (u" GRUPO SISTEMA /EQUIPOS", 4500),
                        (u"SISTEMA /EQUIPOS", 4500),
                        (u"EXISTENCIA", 10000),
                        (u"DESCRIPCIÓN DEL TRABAJO", 10000),
                        (u"CANTIDAD EN LEVANTAMIENTO", 10000),
                    ]
                    row_num = 5
                    for col_num in range(len(columns)):
                            ws.write(row_num, col_num, columns[col_num][0], style_col)
                            ws.col(col_num).width = columns[col_num][1]
                    date_format = xlwt.XFStyle()
                    date_format.num_format_str = 'yyyy/mm/dd'
                    row_num = 6
                    for x in mensual:

                        ws.write(row_num, 0, str(x.usuario_creacion.username), styrow1)
                        ws.write(row_num, 1, str(x.reparacion.gruposistema.descripcion), styrow1)
                        ws.write(row_num, 2, str(x.bien.sistemaequipo), styrow1)
                        ws.write(row_num, 3, str(x.bien.cantidad), styrow1)
                        ws.write(row_num, 4, str(x.descripcion), styrow1)
                        ws.write(row_num, 5, str(x.cantidad), styrow1)
                        row_num += 1
                    # row_num += 3
                    # tecnico=Persona.objects.get(pk=int(persona.id))
                    # ws.write(row_num, 2, '________________________________________________', styrow1)
                    # row_num += 1
                    # ws.write(row_num, 2, str(tecnico.nombre_completo_inverso()), styrow1)
                    # row_num += 1
                    # ws.write(row_num, 2, str(tecnico.mi_cargo()), styrow1)

                    wb.save(response)
                    return response
                except Exception as ex:
                    pass

            if action == 'reportepdfconsolidado':
                try:
                    grupo = int(request.GET['grupo'])
                    bloque = int(request.GET['bloque'])
                    ubicacion = int(request.GET['ubicacion'])
                    tecnico = int(request.GET['tecnico'])

                    if not grupo == 0:
                        gr = ' '
                        gru = HdReparacion.objects.filter(gruposistema=int(grupo), ubicacion__bloque__pk=int(bloque),
                                                          ubicacion__ubicacion__pk=int(ubicacion),
                                                          usuario_creacion=int(persona.usuario.id))
                        for x in gru:
                            gr = x.gruposistema.descripcion
                        data['grupo'] = gr
                    else:
                        data['grupo'] = 'TODOS LOS GRUPOS'
                    if not tecnico == 0:
                        tecnico = Persona.objects.get(status=True, pk=int(request.GET['tecnico'])).usuario.pk

                    else:
                        data['grupo'] = 'TODOS LOS TÉCNICOS'
                    if grupo == 0 and tecnico == 0 and bloque == 0 and ubicacion == 0:
                        mensual = HdDetalle_Reparacion.objects.filter(status=True)
                    elif not grupo == 0 and not tecnico == 0 and not bloque == 0 and not ubicacion == 0:
                        mensual = HdDetalle_Reparacion.objects.filter(reparacion__gruposistema=int(grupo),
                                                                      reparacion__ubicacion__bloque__pk=int(bloque),
                                                                      reparacion__ubicacion__ubicacion__pk=int(
                                                                          ubicacion), usuario_creacion=int(tecnico))
                    elif not grupo == 0 and tecnico == 0 and bloque == 0 and ubicacion == 0:
                        mensual = HdDetalle_Reparacion.objects.filter(reparacion__gruposistema=int(grupo))
                    elif not grupo == 0 and not tecnico == 0 and bloque == 0 and ubicacion == 0:
                        mensual = HdDetalle_Reparacion.objects.filter(reparacion__gruposistema=int(grupo),
                                                                      usuario_creacion=int(tecnico))
                    elif not grupo == 0 and not tecnico == 0 and not bloque == 0 and ubicacion == 0:
                        mensual = HdDetalle_Reparacion.objects.filter(reparacion__gruposistema=int(grupo),
                                                                      reparacion__ubicacion__bloque__pk=int(bloque),
                                                                      usuario_creacion=int(tecnico))
                    elif grupo == 0 and not tecnico == 0 and bloque == 0 and ubicacion == 0:
                        mensual = HdDetalle_Reparacion.objects.filter(usuario_creacion=int(tecnico))
                    elif grupo == 0 and tecnico == 0 and not bloque == 0 and ubicacion == 0:
                        mensual = HdDetalle_Reparacion.objects.filter(reparacion__ubicacion__bloque__pk=int(bloque))
                    elif grupo == 0 and tecnico == 0 and bloque == 0 and not ubicacion == 0:
                        mensual = HdDetalle_Reparacion.objects.filter(
                            reparacion__ubicacion__ubicacion__pk=int(ubicacion))
                    data['reparacion'] = mensual
                    data['tecnico'] =  tecnico = Persona.objects.get(pk=int(persona.id))
                    if  not bloque == 0:
                        data['bloque'] = HdBloqueUbicacion.objects.get(status=True,bloque=bloque,ubicacion=ubicacion)
                    else:
                        data['bloque']='TODOS'

                    data['fechahoy'] = datetime.now().date()
                    return conviert_html_to_pdf(
                        'helpdesk_hdplanificacion/reporteconsolidado.html',
                        {
                            'pagesize': 'A4',
                            'data': data,
                        }
                    )
                except Exception as ex:
                    pass

            if action == 'reportematerialexcel':
                try:

                    mantenimiento=HdMaterialMantenimiento.objects.get(status=True,pk=int(request.GET['id']))
                    materiales=HdMaterialMantenimiento_Material.objects.filter(status=True,materialmantenimiento=int(mantenimiento.pk))
                    responsable=HdMaterialMantenimiento_Responsable.objects.filter(status=True,materialmantenimiento=int(mantenimiento.pk))
                    __author__ = 'Unemi'
                    borders = Borders()
                    borders.left = 1
                    borders.right = 1
                    borders.top = 1
                    borders.bottom = 1
                    styrow = easyxf(
                        'font: name Times New Roman, color-index black, bold off; alignment: horiz centre; borders: left thin, right thin, top thin, bottom thin')
                    styrow1 = easyxf(
                        'font: name Times New Roman, color-index black, bold off; alignment: horiz centre;')
                    style_col = easyxf(
                        'pattern: pattern solid, fore_colour light_turquoise; font: name Times New Roman, color-index black, bold off; alignment: horiz centre; borders: left thin, right thin, top thin, bottom thin')
                    style_sb1 = easyxf('font: name Times New Roman, color-index black, bold on')
                    style_sb = easyxf('font: name Times New Roman, color-index black, bold off')
                    title = easyxf(
                        'font: name Times New Roman, color-index green, bold on , height 350; alignment: horiz centre')
                    title1 = easyxf(
                        'font: name Times New Roman, color-index green, bold on , height 250; alignment: horiz centre')
                    style_date = easyxf(
                        'borders: left thin, right thin, top thin, bottom thin; alignment: horiz centre',
                        num_format_str='yy/mm/dd')
                    style1 = easyxf('borders: left thin, right thin, top thin, bottom thin; alignment: horiz centre',
                                    num_format_str='DD-MM-YYYY')
                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('exp_xls_post_part')
                    ws.write_merge(0, 0, 0, 5, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                    ws.write_merge(1, 1, 0, 5, 'REGISTROS DE MATERIALES', title1)
                    # ws.write_merge(2, 2, 0, 5, '' + ' ' + data['grupo'], title1)
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=Materiales ' + random.randint(1,
                                                                                                                      10000).__str__() + '.xls'


                    columns = [
                        (u"UBICACIÓN", 10000),
                        (u" GRUPO SISTEMA /EQUIPOS", 4500),
                        (u"SISTEMA /EQUIPOS", 4500),
                        (u"TIPO BIEN", 10000),
                        (u"PROCESO", 10000),
                        (u"TIPO MANTENIMIENTO", 10000),
                    ]
                    row_num = 3
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], style_col)
                        ws.col(col_num).width = columns[col_num][1]
                    date_format = xlwt.XFStyle()
                    date_format.num_format_str = 'yyyy/mm/dd'
                    row_num = 4

                    ws.write(row_num, 0, str(mantenimiento.ubicacion.bloque.nombre)+' \n- '+str(mantenimiento.ubicacion.ubicacion.nombre), styrow)
                    ws.write(row_num, 1, str(mantenimiento.gruposistema.descripcion), styrow)
                    ws.write(row_num, 2, str(mantenimiento.bien.sistemaequipo), styrow)
                    ws.write(row_num, 3, str(mantenimiento.get_tipobien_display()), styrow)
                    ws.write(row_num, 4, str(mantenimiento.get_proceso_display()), styrow)
                    ws.write(row_num, 5, str(mantenimiento.get_tipomantenimiento_display()), styrow)
                    row_num += 2
                    ws.write_merge(row_num, row_num, 0, 4, 'LISTADO DE MATERIALES', title1)
                    row_num += 1

                    columns = [
                        (u"MATERIAL", 10000),
                        (u" U. MEDIDA", 4500),
                        (u"CANTIDAD", 4500),
                        (u"PRECIO", 10000),
                        (u"TOTAL", 10000),

                    ]

                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], style_col)
                        ws.col(col_num).width = columns[col_num][1]
                    date_format = xlwt.XFStyle()
                    date_format.num_format_str = 'yyyy/mm/dd'
                    row_num += 1
                    for x in materiales:
                        ws.write(row_num, 0, str(x.material) ,styrow)
                        ws.write(row_num, 1, str(x.unidadmedida.descripcion), styrow)
                        ws.write(row_num, 2, int(x.cantidad), styrow)
                        ws.write(row_num, 3, float(x.precio), styrow)
                        ws.write(row_num, 4, float(x.total), styrow)

                        row_num += 1
                    row_num += 1
                    ws.write_merge(row_num, row_num, 0, 3, 'LISTADO DE RESPONSABLES', title1)
                    row_num += 1

                    columns = [
                        (u"Nº", 10000),
                        (u" CÉDULA", 4500),
                        (u"RESPONSABLE", 8000),
                        (u"CARGO", 10000),


                    ]

                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], style_col)
                        ws.col(col_num).width = columns[col_num][1]
                    date_format = xlwt.XFStyle()
                    date_format.num_format_str = 'yyyy/mm/dd'
                    row_num += 1
                    i=0
                    for x in responsable:
                        i+=1
                        ws.write(row_num, 0, i, styrow)
                        ws.write(row_num, 1, str(x.agente.persona.cedula), styrow)
                        ws.write(row_num, 2, str(x.agente.persona.nombre_completo_inverso()), styrow)
                        ws.write(row_num, 3, str(x.agente.persona.mi_cargo()), styrow)
                        row_num += 1
                    # tecnico=Persona.objects.get(pk=int(persona.id))
                    # ws.write(row_num, 2, '________________________________________________', styrow1)
                    # row_num += 1
                    # ws.write(row_num, 2, str(tecnico.nombre_completo_inverso()), styrow1)
                    # row_num += 1
                    # ws.write(row_num, 2, str(tecnico.mi_cargo()), styrow1)

                    wb.save(response)
                    return response
                except Exception as ex:
                    pass

            if action == 'reportematerialpdf':
                try:
                    mantenimiento = HdMaterialMantenimiento.objects.get(status=True, pk=int(request.GET['id']))
                    materiales = HdMaterialMantenimiento_Material.objects.filter(status=True, materialmantenimiento=int(
                        mantenimiento.pk))
                    responsable = HdMaterialMantenimiento_Responsable.objects.filter(status=True,
                                                                                     materialmantenimiento=int(
                                                                                         mantenimiento.pk))
                    data['mantenimiento'] = mantenimiento
                    data['materiales'] = materiales
                    data['responsable'] = responsable
                    data['fechahoy'] = datetime.now().date()
                    return conviert_html_to_pdf(
                        'helpdesk_hdplanificacion/reportematerialpdf.html',
                        {
                            'pagesize': 'A4',
                            'data': data,
                        }
                    )
                except Exception as ex:
                    pass

            if action == 'cronogramapdfindv':
                try:
                    cronograma = HdCronogramaMantenimiento.objects.get(id=int(request.GET['id']))
                    fechainicio = datetime.strptime(request.GET['fechaInicio'], '%Y-%m-%d').date()
                    fechafin = datetime.strptime(request.GET['fechaFin'], '%Y-%m-%d').date()
                    data['mantenimientosrealizados'] = mantenimientosrealizados = HdDetMantenimientosActivos.objects.filter(
                        status=True, cronograma=cronograma, fecha__gte=fechainicio, fecha__lte=fechafin)
                    if fechainicio != cronograma.desde or fechafin != cronograma.hasta:
                        data['fechacorte'] = [fechainicio,fechafin]

                    data['fechahoy'] = datetime.now().date()
                    data['cronograma'] = cronograma
                    data['contador'] = str(mantenimientosrealizados.count())
                    return conviert_html_to_pdf(
                        'helpdesk_hdplanificacion/cronogramapdfindv.html',
                        {
                            'pagesize': 'A4',
                            'data': data,
                        }
                    )
                except Exception as ex:
                    pass


            if action == 'view':
                data['title'] = u'Registro de incidentes'
                search = None
                ids = None
                if 's' in request.GET:
                    search = request.GET['s']
                    ss = search.split(' ')
                    if len(ss) == 1:
                        incidentes = HdIncidente.objects.filter(asunto__icontains=search, status=True,
                                                                persona_id=persona.id).order_by('estado_id')
                    elif len(ss) == 2:
                        incidentes = HdIncidente.objects.filter(Q(asunto__icontains=ss[0]), Q(asunto__icontains=ss[1]),
                                                                Q(status=True), Q(persona_id=persona.id)).order_by(
                            'estado_id')
                    else:
                        incidentes = HdIncidente.objects.filter(Q(asunto__icontains=ss[0]), Q(asunto__icontains=ss[1]),
                                                                Q(asunto__icontains=ss[1]), Q(status=True),
                                                                Q(persona_id=persona.id)).order_by('estado_id')
                elif 'id' in request.GET:
                    ids = request.GET['id']
                    incidentes = HdIncidente.objects.filter(id=ids, status=True, persona_id=persona.id).order_by(
                        'estado_id')
                else:
                    incidentes = HdIncidente.objects.filter(status=True, persona_id=persona.id).order_by('estado_id')
                paging = MiPaginador(incidentes, 10)
                p = 1
                try:
                    paginasesion = 1
                    if 'paginador' in request.session:
                        paginasesion = int(request.session['paginador'])
                    if 'page' in request.GET:
                        p = int(request.GET['page'])
                    else:
                        p = paginasesion
                    try:
                        page = paging.page(p)
                    except:
                        p = 1
                    page = paging.page(p)
                except:
                    page = paging.page(p)
                request.session['paginador'] = p
                data['paging'] = paging
                data['rangospaging'] = paging.rangos_paginado(p)
                data['page'] = page
                data['search'] = search if search else ""
                data['ids'] = ids if ids else ""
                data['incidentes'] = page.object_list
                data['faltantes'] = 0
                if HdDetEncuestas.objects.filter(encuesta__tipoincidente_id=2, encuesta__activo=True,
                                                 encuesta__status=True, activo=True, status=True).exists():
                    data['faltantes'] = incidentes.filter(tipoincidente_id=2, estado=3, realizoencuesta=False).count()
                data['administrativo'] = persona
                return render(request, "helpdesk_hdusuario/view.html", data)


            if action == 'addpresupuesto':
                try:
                    data['title'] = u'Adicionar  Presupuesto Recursos'
                    form = HdPresupuestoRecursoForm()
                    data['perfilprin'] = perfilprin = request.session['perfilprincipal']
                    data['form'] = form
                    return render(request, "helpdesk_hdplanificacion/addpresupuesto.html", data)
                except Exception as ex:
                    pass
            if action == 'editpresupuesto':
                try:
                    data['title'] = u'Editar   Presupuesto Recursos'
                    data['presupuesto'] = presupuesto = HdPresupuestoRecurso.objects.get(pk=int(request.GET['id']))
                    data['perfilprin'] = perfilprin = request.session['perfilprincipal']
                    form = HdPresupuestoRecursoForm(initial={'presupuestoiva': presupuesto.presupuestoiva,
                                                         'presupuestoreq': presupuesto.presupuestoreq,
                                                         'presupuestototal': presupuesto.presupuestototal,
                                                         'gruposistema': presupuesto.gruposistema,
                                                         'bien': presupuesto.bien,

                                                         })
                    # if presupuesto.gruposistema:
                    #     form.editar(presupuesto.gruposistema)
                    # if presupuesto.bien:
                    #     form.editarbien(presupuesto.bien)
                    data['form'] = form
                    return render(request, "helpdesk_hdplanificacion/editpresupuesto.html", data)
                except Exception as ex:
                    pass
            if action == 'delpresupuesto':
                try:
                    data['title'] = u'Eliminar   Presupuesto Recurso'
                    data['presupuesto'] = HdPresupuestoRecurso.objects.get(pk=int(request.GET['id']))
                    return render(request, "helpdesk_hdplanificacion/deletepresupuesto.html", data)
                except Exception as ex:
                    pass
            if action == 'viewpresupuesto':
                data['title'] = u'Registro  Presupuesto Recurso'
                search = None
                ids = None
                if 's' in request.GET:
                    search = request.GET['s']
                    ss = search.split(' ')
                    # if len(ss) == 1:
                    #     presupuesto = HdPresupuestoRecurso.objects.filter(
                    #                                            status=True,bien__sistemaequipo__icontains=search,gruposistema__descripcion__icontains=search )
                    if len(ss) == 1:
                        presupuesto = HdPresupuestoRecurso.objects.filter(
                                                             Q(bien__sistemaequipo__icontains=ss[0])|Q(gruposistema__descripcion__icontains=ss[0]),
                                                               status=True)

                    else:
                        presupuesto = HdPresupuestoRecurso.objects.filter( Q(bien__sistemaequipo__icontains=ss[0])|Q(gruposistema__descripcion__icontains=ss[0])|Q(bien__sistemaequipo__icontains=ss[1])|Q(gruposistema__descripcion__icontains=ss[1]), Q(status=True))


                elif 'id' in request.GET:
                    ids = request.GET['id']
                    presupuesto = HdPresupuestoRecurso.objects.filter(id=ids, status=True)
                else:
                    presupuesto = HdPresupuestoRecurso.objects.filter(status=True)
                paging = MiPaginador(presupuesto, 10)
                p = 1
                try:
                    paginasesion = 1
                    if 'paginador' in request.session:
                        paginasesion = int(request.session['paginador'])
                    if 'page' in request.GET:
                        p = int(request.GET['page'])
                    else:
                        p = paginasesion
                    try:
                        page = paging.page(p)
                    except:
                        p = 1
                    page = paging.page(p)
                except:
                    page = paging.page(p)
                request.session['paginador'] = p
                data['paging'] = paging
                data['rangospaging'] = paging.rangos_paginado(p)
                data['page'] = page
                data['search'] = search if search else ""
                data['ids'] = ids if ids else ""
                data['presupuesto'] = page.object_list
                data['administrativo'] = persona
                data['fecha'] = datetime.now().date()
                data['grupo'] = HdGrupoSistemaEquipo.objects.filter(status=True)
                data['anio'] = AnioEjercicio.objects.filter(status=True).order_by('-anioejercicio')
                return render(request, "helpdesk_hdplanificacion/viewpresupuesto.html", data)

            elif action == 'viewtipobien':
                try:
                    data['title'] = u'Pieza partes'
                    url_vars = ''
                    request.session['viewactivo'] = 6
                    tipobien = HdGruposCategoria.objects.filter(status=True)
                    search = None
                    if 's' in request.GET:
                        search = request.GET['s']
                        url_vars += "&s={}".format(search)
                        tipobien = tipobien.filter(descripcion__icontains=search)
                    paging = MiPaginador(tipobien,20)
                    p=1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p=1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ''
                    data['tipobien'] = page.object_list
                    data["url_vars"] = url_vars
                    return render(request,"helpdesk_hdplanificacion/viewtipobien.html", data)

                except Exception as ex:
                    pass

            elif action == 'addtipobien':
                try:
                    data['title'] = u'Adicionar tipo bien'
                    form = TipoBienForm()
                    data['form'] = form
                    template = get_template("helpdesk_hdplanificacion/modal/addtipobien.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, 'message': str(ex)})

            elif action == 'edittipobien':
                try:
                    data['title'] = u'Adicionar tipo bien'
                    data['id'] = id = int(encrypt(request.GET['id']))
                    filtro = HdGruposCategoria.objects.get(pk = id)
                    form = TipoBienForm(initial=model_to_dict(filtro))
                    data['form'] = form
                    template = get_template("helpdesk_hdplanificacion/modal/addtipobien.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, 'message': str(ex)})

            elif action == 'viewtlimpieza':
                try:
                    data['title'] = u'Tareas de mantenimiento'
                    url_vars = ''
                    request.session['viewactivo'] = 7
                    gruposelect = request.GET.get('tipo','0')
                    tlimpieza = HdMantenimientoGruCategoria.objects.filter(status=True).order_by('grupocategoria_id', 'descripcion')
                    search = None
                    tipo = None
                    grupos = HdGruposCategoria.objects.filter(status=True)
                    listagrupo = []
                    if gruposelect != '0':
                        data['gruposelect'] = gruposelect = int(encrypt(request.GET['tipo']))
                        url_vars += "&tipo={}".format(gruposelect)
                        if gruposelect > 0:
                            listagrupo.append(gruposelect)
                        else:
                            listagrupo = grupos.values_list('id')
                    else:
                        listagrupo = grupos.values_list('id')
                    if 's' in request.GET:
                        search = request.GET['s']
                        url_vars += "&s={}".format(search)
                        tlimpieza = tlimpieza.filter(Q(descripcion__icontains=search)|Q(grupocategoria__descripcion__icontains=search)).order_by('grupocategoria_id', 'descripcion')

                    if gruposelect != '0':
                        tlimpieza = tlimpieza.filter(grupocategoria_id__in = listagrupo).order_by('grupocategoria_id', 'descripcion')
                    paging = MiPaginador(tlimpieza,20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data['tipo'] = tipo if tipo else ""
                    data['tareasmantenimiento'] = page.object_list
                    data['tipobien'] = grupos
                    return render(request,"helpdesk_hdplanificacion/viewtareaslimp.html",data)
                except Exception as ex:
                    pass

            elif action == 'addtarealimp':
                try:
                    data['title'] = u'Adicionar tarea de limpieza'
                    form = TareasMantenimientoForm()
                    data['form'] = form
                    template = get_template("helpdesk_hdplanificacion/modal/formtareaslimpieza.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, 'message': str(ex)})

            elif action == 'edittarealimp':
                try:
                    data['title'] = u'Editar tarea de limpieza'
                    data['id'] = id = int(encrypt(request.GET['id']))
                    filtro = HdMantenimientoGruCategoria.objects.get(pk = id)
                    form = TareasMantenimientoForm(initial={
                        'tipobien':filtro.grupocategoria,
                        'descripcion':filtro.descripcion
                    })
                    data['form'] = form
                    template = get_template("helpdesk_hdplanificacion/modal/formtareaslimpieza.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, 'message': str(ex)})

            elif action == 'viewdanios':
                try:
                    data['title'] = u'Lista de daños'
                    request.session['viewactivo'] = 8
                    url_vars = ''
                    gruposelect = request.GET.get('tipo', '0')
                    tdanios = HdMantenimientoGruDanios.objects.filter(status=True).order_by('grupocategoria_id', 'descripcion')
                    search = None
                    tipo = None
                    grupos = HdGruposCategoria.objects.filter(status=True)
                    listagrupo = []
                    if gruposelect != '0':
                        data['gruposelect'] = gruposelect = int(encrypt(request.GET['tipo']))
                        url_vars += "&tipo={}".format(gruposelect)
                        if gruposelect > 0:
                            listagrupo.append(gruposelect)
                        else:
                            listagrupo = grupos.values_list('id')
                    else:
                        listagrupo = grupos.values_list('id')
                    if 's' in request.GET:
                        search = request.GET['s']
                        url_vars += "&s={}".format(search)
                        tdanios = tdanios.filter(Q(descripcion__icontains=search)|Q(grupocategoria__descripcion__icontains=search)).order_by('grupocategoria_id', 'descripcion')

                    if gruposelect != '0':
                        tdanios = tdanios.filter(grupocategoria_id__in = listagrupo).order_by('grupocategoria_id', 'descripcion')
                    paging = MiPaginador(tdanios,20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data['tipo'] = tipo if tipo else ""
                    data['tareasmantenimiento'] = page.object_list
                    data['tipobien'] = grupos
                    return render(request,"helpdesk_hdplanificacion/viewdanios.html",data)
                except Exception as ex:
                    pass

            elif action == 'adddanios':
                try:
                    data['title'] = u'Adicionar daño'
                    form = HdMantenimientoGruDaniosForm()
                    data['form'] = form
                    template = get_template("helpdesk_hdplanificacion/modal/formtareaslimpieza.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, 'message': str(ex)})

            elif action == 'editdanios':
                try:
                    data['title'] = u'Editar daño'
                    data['id'] = id = int(encrypt(request.GET['id']))
                    filtro = HdMantenimientoGruDanios.objects.get(pk = id)
                    form = HdMantenimientoGruDaniosForm(initial={
                        'tipobien':filtro.grupocategoria,
                        'descripcion':filtro.descripcion
                    })
                    data['form'] = form
                    template = get_template("helpdesk_hdplanificacion/modal/formtareaslimpieza.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, 'message': str(ex)})

            elif action == 'viewpiezaparte':
                try:
                    request.session['viewactivo'] = 9
                    data['title'] = u'Lista de pieza-partes'
                    url_vars = ''
                    gruposelect = request.GET.get('tipo', '0')
                    tpiezapartes = HdPiezaPartes.objects.filter(status=True).order_by('grupocategoria_id', 'descripcion')
                    search = None
                    tipo = None
                    grupos = HdGruposCategoria.objects.filter(status=True)
                    listagrupo = []
                    if gruposelect != '0':
                        data['gruposelect'] = gruposelect = int(encrypt(request.GET['tipo']))
                        url_vars += "&tipo={}".format(gruposelect)
                        if gruposelect > 0:
                            listagrupo.append(gruposelect)
                        else:
                            listagrupo = grupos.values_list('id')
                    else:
                        listagrupo = grupos.values_list('id')
                    if 's' in request.GET:
                        search = request.GET['s']
                        url_vars += "&s={}".format(search)
                        tpiezapartes = tpiezapartes.filter(Q(descripcion__icontains=search)|Q(grupocategoria__descripcion__icontains=search)).order_by('grupocategoria_id', 'descripcion')

                    if gruposelect != '0':
                        tpiezapartes = tpiezapartes.filter(grupocategoria_id__in = listagrupo).order_by('grupocategoria_id', 'descripcion')
                    paging = MiPaginador(tpiezapartes,20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data['tipo'] = tipo if tipo else ""
                    data['piezapartes'] = page.object_list
                    data['tipobien'] = grupos
                    return render(request, "helpdesk_hdplanificacion/viewpiezapartes.html", data)
                except Exception as ex:
                    pass

            elif action == 'addpiezaparte':
                try:
                    data['title'] = u'Adicionar pieza-parte'
                    form = HdPiezaPartesForm()
                    data['form'] = form
                    template = get_template("helpdesk_hdplanificacion/modal/formpiezaparte.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, 'message': str(ex)})

            elif action == 'editpiezaparte':
                try:
                    data['title'] = u'Editar pieza-parte'
                    data['id'] = id = int(encrypt(request.GET['id']))
                    filtro = HdPiezaPartes.objects.get(pk = id)
                    form = HdPiezaPartesForm(initial={
                        'tipobien':filtro.grupocategoria,
                        'descripcion':filtro.descripcion,
                        'estado':filtro.estado,
                        'imagen':filtro.imagen
                    })
                    data['form'] = form
                    template = get_template("helpdesk_hdplanificacion/modal/formpiezaparte.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, 'message': str(ex)})

            elif action == 'addresponsable':
                try:
                    data['activo'] = activo = ActivoTecnologico.objects.get(pk=int(request.GET['id']), status=True)
                    form = ResponsableActivoTraspasoForm(initial={
                        'responsableactual':activo.activotecnologico.responsable,
                        'activo':activo
                    })
                    data['form2'] = form
                    template = get_template('helpdesk_hdplanificacion/modal/formasignarresponsable.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({'result':False,'mensaje':u'Error al cargar los datos!'})

            elif action == 'informe_detallemantenimiento':
                try:
                    data['mantenimiento'] = mantenimiento = HdDetMantenimientosActivos.objects.get(pk=encrypt(request.GET['id']))
                    data['tareasmantenimiento'] = HdMantenimientoGruCategoria.objects.filter(grupocategoria=mantenimiento.tipoactivoc, status=True, activo=True)
                    data['tareasactivo'] = HdTareasActivosPreventivos.objects.values_list('grupos_id',flat=True).filter(mantenimiento=mantenimiento, status=True)
                    data['ppmantenimiento'] = HdPiezaPartes.objects.filter(grupocategoria=mantenimiento.tipoactivoc,status=True)
                    data['piezaparteactivo'] = HdPiezaParteActivosPreventivos.objects.filter(mantenimiento=mantenimiento, status=True)
                    data['daniomantenimiento'] = HdMantenimientoGruDanios.objects.filter(grupocategoria=mantenimiento.tipoactivoc, status=True, activo=True)
                    data['danioactivo'] = HdTareasActivosPreventivosDanios.objects.filter(mantenimiento=mantenimiento,status=True)
                    data['estdan'] = ESTADO_DANIO
                    return conviert_html_to_pdf('helpdesk_hdplanificacion/informes/reportemantenimientotecnologico.html',{'pagesize': 'A4', 'data': data})
                except Exception as ex:
                    pass

            elif action == 'reportecronogramaindv':
                try:
                    id = int(encrypt(request.GET['id']))
                    cronograma = HdCronogramaMantenimiento.objects.get(pk=id)
                    data['cronograma'] = cronograma
                    data['action'] = 'cronogramapdfindv'
                    data['request'] = request
                    template = get_template('helpdesk_hdplanificacion/modal/reportecrono.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({'result': False, 'mensaje': u'Error al cargar los datos!'})

            return HttpResponseRedirect(request.path)

        else:
            data['title'] = u'Cronograma Mantenimientos'
            search = None
            ids = None
            id_estado = 0
            idgrupo = 0
            cronograma = HdCronogramaMantenimiento.objects.filter( status=True).order_by('fecha_creacion')
            if 's' in request.GET:
                search = request.GET['s']
                ss = search.split(' ')
                if len(ss) == 1:
                    cronograma = HdCronogramaMantenimiento.objects.filter(gruposistema__descripcion__icontains=search, status=True).order_by('fecha_creacion')
                elif len(ss) == 2:
                    cronograma = HdCronogramaMantenimiento.objects.filter(Q(gruposistema__descripcion__icontains=ss[0])|Q(gruposistema__descripcion__icontains=ss[1]), Q(status=True)).order_by('fecha_creacion')
                else:
                    cronograma = HdCronogramaMantenimiento.objects.filter(Q(gruposistema__descripcion__icontains=ss[0]) |Q(gruposistema__descripcion__icontains=ss[1]), Q(status=True)).order_by('fecha_creacion')

            if 'idg' in request.GET and not  'id_estado' in request.GET:
                idgrupo = int(request.GET['idg'])
                if idgrupo != 0:
                    cronograma = HdCronogramaMantenimiento.objects.filter( status=True,gruposistema__id=idgrupo).order_by('fecha_creacion')
            elif 'id_estado' in request.GET and not 'idg' in request.GET:
                id_estado = int(request.GET['id_estado'])
                if id_estado != 0:
                    cronograma = HdCronogramaMantenimiento.objects.filter(status=True,tipomantenimiento=id_estado).order_by('fecha_creacion')
            elif 'id_estado' in request.GET and 'idg' in request.GET:
                id_estado = int(request.GET['id_estado'])
                idgrupo = int(request.GET['idg'])
                if id_estado != 0 and  idgrupo != 0:
                    cronograma = HdCronogramaMantenimiento.objects.filter(status=True,tipomantenimiento=id_estado,gruposistema__id=idgrupo).order_by('fecha_creacion')
            if 'id' in request.GET:
                ids = request.GET['id']
                cronograma = HdCronogramaMantenimiento.objects.filter(id=ids, status=True).order_by('fecha_creacion')
            # else:
            #     cronograma = HdCronogramaMantenimiento.objects.filter(status=True).order_by('fecha_creacion')
            paging = MiPaginador(cronograma, 10)
            p = 1
            try:
                paginasesion = 1
                if 'paginador' in request.session:
                    paginasesion = int(request.session['paginador'])
                if 'page' in request.GET:
                    p = int(request.GET['page'])
                else:
                    p = paginasesion
                try:
                    page = paging.page(p)
                except:
                    p = 1
                page = paging.page(p)
            except:
                page = paging.page(p)
            request.session['paginador'] = p
            request.session['viewactivo'] = 1
            data['paging'] = paging
            data['rangospaging'] = paging.rangos_paginado(p)
            data['page'] = page
            data['search'] = search if search else ""
            data['ids'] = ids if ids else ""
            data['cronograma'] = page.object_list
            data['administrativo'] = persona
            data['fecha'] = datetime.now().date()
            data['grupo'] = HdGrupoSistemaEquipo.objects.filter(status=True)
            data['meses'] = MONTH_CHOICES
            data['anio'] = AnioEjercicio.objects.filter(status=True).order_by('-anioejercicio')
            data['tipo'] = TIPO_MANTENIMIENTO
            data['estadoid'] = id_estado

            # data['estadoOR'] = ESTADO_ORDEN_TRABAJO
            data['idgrupo'] = idgrupo
            return render(request, "helpdesk_hdplanificacion/view.html", data)


