import calendar
import json
import sys
from datetime import datetime, timedelta, date

from docutils.nodes import status
from openpyxl import workbook as openxl
from openpyxl.chart import ScatterChart, Reference, Series,PieChart, BarChart
from openpyxl.styles import Font as openxlFont
from openpyxl.styles.alignment import Alignment as alin

from django.contrib.auth.decorators import login_required
from django.forms import model_to_dict
from django.http import HttpResponseRedirect, JsonResponse, HttpResponse
from django.shortcuts import render
from django.template.loader import get_template
from xlwt import easyxf, XFStyle, Workbook
from collections import defaultdict
from balcon.models import RespuestaEncuestaSatisfaccion, Proceso, CategoriaEncuesta, EncuestaProceso
from decorators import secure_module, last_access
from inno.models import BecaPeriodoResumen, AsistenciaMatriculadoPeriodo
from poli.models import ActividadPolideportivo
from sagest.commonviews import total_efectivo_dia, cantidad_facturas_dia, cantidad_cheques_dia, total_cheque_dia, cantidad_tarjetas_dia, total_tarjeta_dia, cantidad_depositos_dia, total_deposito_dia, cantidad_transferencias_dia, \
    total_transferencia_dia, total_dia, total_efectivo_mes, total_efectivo_rango, cantidad_cheques_rango, total_cheque_rango, cantidad_tarjetas_rango, \
    total_tarjeta_rango, cantidad_depositos_rango, total_deposito_rango, cantidad_transferencias_rango, total_transferencia_rango, cantidad_facturas_rango, total_rango
from sagest.funciones import encrypt_id
from sga.commonviews import adduserdata, total_matriculados, total_matriculados_hombres, total_matriculados_mujeres, cantidad_matriculados_discapacidad, cantidad_matriculados_beca, total_matriculados_menor_30, \
    total_matriculados_31_40, total_matriculados_41_50, total_matriculados_51_60, total_matriculados_mayor_61, porciento_matriculados_discapacidad, porciento_matriculados_beca, total_prematriculados, \
    total_ventas_edades, total_ventas_prov_sexo, traerNotificaciones
from sga.excelbackground import  reporte_matriculados_asistencia_background
from sagest.commonviews import cantidad_total_deudores, valor_total_deudores_activos_30dias, valor_total_deudores_activos_31_90dias, valor_total_deudores_activos_mas_90dias, \
    valor_deudores_activos_total, valor_total_apagar_activos_30dias, valor_total_apagar_activos_31_90dias, \
    valor_total_apagar_activos_mas_90dias, cantidad_total_apagar, valor_apagar_activos_total, valor_deudas_activos_total

from sga.funciones import convertir_fecha, convertir_fecha_invertida, null_to_decimal, cantidad_evaluacion_docente, cantidad_evaluacion_auto, \
    evaluo_director, evaluo_coordinador
from sga.models import Coordinacion, Carrera, Provincia, Canton, PreMatricula, Matricula, Sesion, PreInscripcionBeca, \
    BecaTipo, BecaSolicitud, BecaAsignacion, BecaSolicitudRecorrido, ESTADO_SOLICITUD_BECAS, ESTADO_ACEPTACION_BECA, \
    ESTADO_ASIGNACION_BECA, Periodo, Malla, Modalidad, NivelMalla, Nivel, Pais, MateriaAsignada, Notificacion, ProfesorMateria
from sagest.models import SesionCaja, Rubro
from sga.templatetags.sga_extras import traducir_mes, transformar_mes, diaenletra, encrypt
from socioecon.models import GrupoSocioEconomico, cantidad_gruposocioeconomico_carrera, \
    cantidad_gruposocioeconomico_coordinacion, TipoHogar, cantidad_tipo_hogar_coordinacion, cantidad_tipo_hogar_carrera, \
    NivelEstudio, cantidad_nivel_educacion_jefehogar_carrera, cantidad_nivel_educacion_jefehogar_coordinacion, \
    cantidad_sidependientes_carrera, cantidad_sidependientes_coordinacion, cantidad_nodependientes_carrera, \
    cantidad_nodependientes_coordinacion, cantidad_sicabezasf_carrera, cantidad_nocabezasf_carrera, \
    cantidad_sicabezasf_coordinacion, cantidad_nocabezasf_coordinacion
from django.db import transaction
from posgrado.models import *
from posgrado.models import InscripcionCohorte, VentasProgramaMaestria
from django.db.models import Q, F, Value as V, Count, OuterRef, Subquery
from django.db.models.functions import ExtractYear, ExtractMonth, ExtractDay
from django.db.models.functions import Concat
import xlsxwriter
import io
from django.db.models import F, Sum
from django.db.models import FloatField
from django.db.models.functions import Coalesce

#preuba WA
import openpyxl
from openpyxl import load_workbook
from xlwt import *
from django.views.generic import TemplateView
from django.db import connections

def cincoacien(valor):
    return round((valor * 100 / 5), 2)

def fecha_vencida(obj, tipo):
    from posgrado.models import DetalleFechasEvalDirMateria
    try:
        estado = 'NO CONFIGURADO'
        if tipo == 1:
            if obj.inicioeval and obj.fineval:
                if datetime.now().date() > obj.fineval:
                    estado = 'FINALIZADA'
                else:
                    estado = 'EN CURSO'
        elif tipo == 2:
            if obj.inicioevalauto and obj.finevalauto:
                if datetime.now().date() > obj.finevalauto:
                    estado = 'FINALIZADA'
                else:
                    estado = 'EN CURSO'
        elif tipo == 3:
            if DetalleFechasEvalDirMateria.objects.filter(status=True, materia=obj).exists():
                eDetalle = DetalleFechasEvalDirMateria.objects.filter(status=True, materia=obj).first()
                if datetime.now().date() > eDetalle.fin:
                    estado = 'FINALIZADA'
                else:
                    estado = 'EN CURSO'
        return estado
    except Exception as ex:
        pass

def object_dir(obj):
    eDetalle = None
    if DetalleFechasEvalDirMateria.objects.filter(status=True, materia=obj.materia).exists():
        eDetalle = DetalleFechasEvalDirMateria.objects.filter(status=True, materia=obj.materia).first()
    return eDetalle

def object_proces(obj):
    eDetalle = None
    if DetalleResultadosEvaluacionPosgrado.objects.filter(status=True, materia=obj.materia).exists():
        eDetalle = DetalleResultadosEvaluacionPosgrado.objects.filter(status=True, materia=obj.materia).first()
    return eDetalle

@login_required(redirect_field_name='ret', login_url='/loginsga')
@secure_module
@last_access
@transaction.atomic()
def view(request):
    global ex
    data = {}
    adduserdata(request, data)
    persona = request.session['persona']
    data['persona'] = persona
    if request.method == 'POST':
        action = request.POST['action']
        periodo = request.session['periodo']

        if action == 'loadDataTableMatriculadosTipo':
            try:
                txt_filter = request.POST['sSearch'] if request.POST['sSearch'] else ''
                limit = int(request.POST['iDisplayLength']) if request.POST['iDisplayLength'] else 25
                offset = int(request.POST['iDisplayStart']) if request.POST['iDisplayStart'] else 0
                tCount = 0
                tipo_id = request.POST.get('tipomatricula_id')
                retiradomatricula = request.POST.get('retiradomatricula')
                matriculados = Matricula.objects.filter(status=True,
                                                        nivel__periodo_id=periodo.id,
                                                        matriculagruposocioeconomico__tipomatricula=tipo_id,
                                                        matriculagruposocioeconomico__status=True
                                                        ).distinct()
                if retiradomatricula is not None:
                    retiradomatricula = json.loads(retiradomatricula)
                    matriculados = matriculados.filter(retiradomatricula=retiradomatricula)
                if txt_filter:
                    search = txt_filter.strip()
                    ss = search.split(' ')
                    if len(ss) == 1:
                        matriculados = matriculados.filter(Q(inscripcion__persona__cedula__icontains=search) |
                                                   Q(inscripcion__persona__pasaporte__icontains=search) |
                                                   Q(inscripcion__persona__ruc__icontains=search) |
                                                   Q(inscripcion__persona__nombres__icontains=search) |
                                                   Q(inscripcion__persona__apellido1__icontains=search) |
                                                   Q(inscripcion__persona__apellido2__icontains=search)
                                                   )
                    else:
                        matriculados = matriculados.filter(Q(inscripcion__persona__nombres__icontains=search) |
                                                   Q(inscripcion__persona__apellido1__icontains=ss[0]))
                tCount = matriculados.count()
                if offset == 0:
                    rows = matriculados[offset:limit]
                else:
                    rows = matriculados[offset:offset + limit]
                aaData = []
                for row in rows:
                    aaData.append([row.inscripcion.persona.__str__(),
                                   row.inscripcion.persona.documento(),
                                   row.inscripcion.persona.sexo.__str__() if row.inscripcion.persona.sexo is not None else 'S/N',
                                   row.inscripcion.carrera.__str__(),
                                   row.inscripcion.mi_nivel().__str__(),
                                   {"id": row.id},

                                   ])
                return JsonResponse({"result": "ok", "data": aaData, "iTotalRecords": tCount, "iTotalDisplayRecords": tCount})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al cargar los datos. %s" % ex.__str__(), "data": [], "iTotalRecords": 0, "iTotalDisplayRecords": 0})

        if action == 'loadDataTablePreseleccionadosTipoBeca':
            try:
                txt_filter = request.POST['sSearch'] if request.POST['sSearch'] else ''
                limit = int(request.POST['iDisplayLength']) if request.POST['iDisplayLength'] else 25
                offset = int(request.POST['iDisplayStart']) if request.POST['iDisplayStart'] else 0
                tCount = 0
                becatipo_id = request.POST.get('becatipo_id')
                matriculados = PreInscripcionBeca.objects.filter(status=True,
                                                        periodo_id=periodo.id,
                                                        becatipo=becatipo_id)
                if txt_filter:
                    search = txt_filter.strip()
                    ss = search.split(' ')
                    if len(ss) == 1:
                        matriculados = matriculados.filter(Q(inscripcion__persona__cedula__icontains=search) |
                                                   Q(inscripcion__persona__pasaporte__icontains=search) |
                                                   Q(inscripcion__persona__ruc__icontains=search) |
                                                   Q(inscripcion__persona__nombres__icontains=search) |
                                                   Q(inscripcion__persona__apellido1__icontains=search) |
                                                   Q(inscripcion__persona__apellido2__icontains=search)
                                                   )
                    else:
                        matriculados = matriculados.filter(Q(inscripcion__persona__nombres__icontains=search) |
                                                   Q(inscripcion__persona__apellido1__icontains=ss[0]))
                tCount = matriculados.count()
                if offset == 0:
                    rows = matriculados[offset:limit]
                else:
                    rows = matriculados[offset:offset + limit]
                aaData = []
                for row in rows:
                    aaData.append([row.inscripcion.persona.__str__(),
                                   row.inscripcion.persona.documento(),
                                   row.inscripcion.persona.sexo.__str__() if row.inscripcion.persona.sexo is not None else 'S/N',
                                   row.inscripcion.carrera.__str__(),
                                   row.inscripcion.mi_nivel().__str__(),
                                   {"id": row.id},

                                   ])
                return JsonResponse({"result": "ok", "data": aaData, "iTotalRecords": tCount, "iTotalDisplayRecords": tCount})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al cargar los datos. %s" % ex.__str__(), "data": [], "iTotalRecords": 0, "iTotalDisplayRecords": 0})

        if action == 'loadDataTableAdjudicadosTipoBeca':
            try:
                txt_filter = request.POST['sSearch'] if request.POST['sSearch'] else ''
                limit = int(request.POST['iDisplayLength']) if request.POST['iDisplayLength'] else 25
                offset = int(request.POST['iDisplayStart']) if request.POST['iDisplayStart'] else 0
                tCount = 0
                becatipo_id = request.POST.get('becatipo_id')
                becaaceptada_id = request.POST.get('becaaceptada_id')
                becaasignada_id = request.POST.get('becaasignada_id')
                estado_id = request.POST.get('estado_id')
                control_accion = request.POST.get('control_acciones')
                matriculados = BecaSolicitud.objects.filter(status=True,
                                                            periodo_id=periodo.id,
                                                            becatipo=becatipo_id)
                if becaaceptada_id is not None:
                    matriculados = matriculados.filter(becaaceptada=becaaceptada_id) if becaaceptada_id != '0' else matriculados.filter(becaaceptada__isnull=False)

                if becaasignada_id is not None:
                    matriculados = matriculados.filter(becaasignada=becaasignada_id) if becaasignada_id != '0' else matriculados.filter(becaasignada__isnull=False)

                if estado_id is not None:
                    subquery = BecaSolicitudRecorrido.objects.filter(solicitud=OuterRef('pk'), status=True).order_by('-fecha_creacion')
                    matriculados = matriculados.annotate(estado_ultimo=Subquery(subquery.values('estado')[:1]))
                    matriculados = matriculados.filter(estado_ultimo=estado_id) if estado_id != '0' else matriculados

                if txt_filter:
                    search = txt_filter.strip()
                    ss = search.split(' ')
                    if len(ss) == 1:
                        matriculados = matriculados.filter(Q(inscripcion__persona__cedula__icontains=search) |
                                                   Q(inscripcion__persona__pasaporte__icontains=search) |
                                                   Q(inscripcion__persona__ruc__icontains=search) |
                                                   Q(inscripcion__persona__nombres__icontains=search) |
                                                   Q(inscripcion__persona__apellido1__icontains=search) |
                                                   Q(inscripcion__persona__apellido2__icontains=search)
                                                   )
                    else:
                        matriculados = matriculados.filter(Q(inscripcion__persona__nombres__icontains=search) |
                                                   Q(inscripcion__persona__apellido1__icontains=ss[0]))
                tCount = matriculados.count()
                if offset == 0:
                    rows = matriculados[offset:limit]
                else:
                    rows = matriculados[offset:offset + limit]
                aaData = []
                control_accion = json.loads(control_accion) if not control_accion in [None, ''] else False

                for row in rows:
                    acciones = ''
                    if control_accion:
                        acciones = f"""
                                    <div class="table-controls">
                                        <div class="btn-group">
                                            <input type="hidden" class="dt-col-option" value="{row.id}"/>
                                            <a class="btn btn-mini dropdown-toggle" data-toggle="dropdown" href="javascript:;">Acciones<span class="caret"></span>
                                            </a>
                                            <ul class="dropdown-menu pull-right" style="text-align: left">
                                                <li><a href="javascript:;" class="dt-action-view-route"><i class="fa fa-eye" aria-hidden="true"></i> Ver Recorrido</a></li>
                                            </ul>
                                            <input type="hidden" class="dt-col-data-name" value="{row.__str__()}"/>
                                        </div>
                                    </div>
                                    """
                    aaData.append([row.inscripcion.persona.__str__(),
                                   row.inscripcion.persona.documento(),
                                   row.inscripcion.persona.sexo.__str__() if row.inscripcion.persona.sexo is not None else 'S/N',
                                   row.inscripcion.carrera.__str__(),
                                   row.inscripcion.mi_nivel().__str__(),
                                   acciones,
                                   {"id": row.id},

                                   ])
                return JsonResponse({"result": "ok", "data": aaData, "iTotalRecords": tCount, "iTotalDisplayRecords": tCount})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al cargar los datos. %s" % ex.__str__(), "data": [], "iTotalRecords": 0, "iTotalDisplayRecords": 0})

        if action == 'loadRecorrido':
            try:
                id = request.POST.get('id')
                data['solicitud'] = solicitud = BecaSolicitud.objects.filter(pk=id).first()
                if solicitud is None:
                    raise NameError('No existe Solicitud de Beca')
                data['recorridos'] = solicitud.becasolicitudrecorrido_set.filter(status=True)
                template = get_template("estadisticas/partials/recorrido.html")
                json_content = template.render(data, request=request)
                return JsonResponse({"result": "ok", 'html': json_content})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al cargar los datos."})

        if action == 'saveBecaPeriodoResumen':
            try:
                value = int(request.POST.get('value'))
                id = int(encrypt(request.POST.get('id')))
                op = int(request.POST.get('op'))
                becaperiodoresumen = BecaPeriodoResumen.objects.get(pk=id)
                porcentaje = None
                if op == 1:
                    becaperiodoresumen.matriculados = value
                elif op == 2:
                    becaperiodoresumen.matriculados_regulares = value
                elif op == 3:
                    becaperiodoresumen.preseleccionados_becas = value
                elif op == 4:
                    becaperiodoresumen.preseleccionados_becasaceptados = value
                elif op == 5:
                    becaperiodoresumen.preseleccionados_becasadjudicados = value
                elif op == 6:
                    becaperiodoresumen.preseleccionados_becaspagados = value
                becaperiodoresumen.save(request)
                model_data = model_to_dict(becaperiodoresumen, exclude=['fecha_cohorte', 'archivo'])
                model_data['porcentaje_preseleccionados'] = becaperiodoresumen.porcentaje_de_cumplimiento_preseleccionados()
                model_data['porcentaje_aceptados'] = becaperiodoresumen.porcentaje_de_cumplimiento_preseleccionados_aceptados()
                model_data['porcentaje_adjudicados'] = becaperiodoresumen.porcentaje_de_cumplimiento_preseleccionados_adjudicadas()
                model_data['porcentaje_pagados'] = becaperiodoresumen.porcentaje_de_cumplimiento_preseleccionados_pagados()
                return JsonResponse({"result": "ok", 'model_data': model_data, 'value': value, 'id': id})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al cargar los datos."})

        elif action == 'selectcohorte':
            try:
                if 'id' in request.POST:
                    lista = []
                    cohortes = CohorteMaestria.objects.filter(maestriaadmision_id=int(request.POST['id']), status=True).order_by('-id')
                    for cohorte in cohortes:
                        lista.append([cohorte.id, cohorte.descripcion])
                    return JsonResponse({"result": "ok", 'lista': lista})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al consulatar los datos."})

        elif action == 'selectprovincia':
            try:
                if 'id' in request.POST:
                    lista = []
                    provincias = Provincia.objects.filter(pais_id=int(request.POST['id']), status=True).order_by('nombre')
                    for provincia in provincias:
                        lista.append([provincia.id, provincia.nombre])
                    return JsonResponse({"result": "ok", 'lista': lista})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al consulatar los datos."})

        elif action == 'selectcanton':
            try:
                if 'id' in request.POST:
                    lista = []
                    cantones = Canton.objects.filter(provincia_id=int(request.POST['id']), status=True).order_by('nombre')
                    for canton in cantones:
                        lista.append([canton.id, canton.nombre])
                    return JsonResponse({"result": "ok", 'lista': lista})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al consulatar los datos."})

        return JsonResponse({"result": "bad", "mensaje": u"Solicitud Incorrecta."})
    else:
        if 'action' in request.GET:
            data['action']= action = request.GET['action']
            periodo = request.session['periodo']

            if action == 'tbl_gpo_carrera':
                try:
                    # CONFORMACION DE TABLA DE GRUPOS SOCIOECONOMICOS POR CARRERAS
                    data['title'] = u'Grupos Socioeconomicos por carreras'
                    carreras = Carrera.objects.filter(activa=True)
                    lista_carreras_grupos = []
                    for c in carreras:
                        lista_grupos = []
                        for grupo in GrupoSocioEconomico.objects.all():
                            lista_grupos.append(cantidad_gruposocioeconomico_carrera(grupo, c))
                        lista_carreras_grupos.append((c.alias, lista_grupos))
                    data['carreras'] = carreras
                    data['lista_carreras_grupos'] = lista_carreras_grupos
                    data['grupos_socioeconomicos'] = GrupoSocioEconomico.objects.all()
                    return render(request, "estadisticas/tbl_gpo_carrera.html", data)
                except Exception as ex:
                    pass

            elif action == 'tbl_gpo_coordinacion':
                try:
                    # CONFORMACION DE TABLA DE GRUPOS SOCIOECONOMICOS POR COORDINACIONES
                    data['title'] = u'Grupos Socioeconomicos por coordinaciones'
                    coordinaciones = Coordinacion.objects.all().order_by('id')
                    lista_coordinaciones_grupos = []
                    for c in coordinaciones:
                        lista_grupos = []
                        for grupo in GrupoSocioEconomico.objects.all():
                            lista_grupos.append(cantidad_gruposocioeconomico_coordinacion(grupo, c))
                        lista_coordinaciones_grupos.append((c.nombre, lista_grupos))
                    data['coordinaciones'] = coordinaciones
                    data['lista_coordinaciones_grupos'] = lista_coordinaciones_grupos
                    data['grupos_socioeconomicos'] = GrupoSocioEconomico.objects.all()
                    return render(request, "estadisticas/tbl_gpo_coord.html", data)
                except Exception as ex:
                    pass

            elif action == 'graph_tipohogar':
                try:
                    # CONFORMACION DE GRAFICOS PARA MOSTRAR EL TIPO DE HOGAR DE ESTUDIANTES
                    data['title'] = u'Tipos de hogares de lo estudiantes'
                    coordinaciones = Coordinacion.objects.all().order_by('id')
                    carreras = Carrera.objects.all()
                    lista_carreras_tipo_hogar = []
                    for c in carreras:
                        lista_tipohogar = []
                        for th in TipoHogar.objects.all():
                            lista_tipohogar.append(cantidad_tipo_hogar_carrera(th, c))
                        lista_carreras_tipo_hogar.append((c.alias, lista_tipohogar))
                    data['carreras'] = carreras
                    data['lista_carreras_tipo_hogar'] = lista_carreras_tipo_hogar
                    lista_coordinaciones_tipo_hogar = []
                    for c in coordinaciones:
                        lista_tipohogar = []
                        for th in TipoHogar.objects.all():
                            lista_tipohogar.append(cantidad_tipo_hogar_coordinacion(th, c))
                        lista_coordinaciones_tipo_hogar.append((c.nombre, lista_tipohogar))
                    data['coordinaciones'] = coordinaciones
                    data['lista_coordinaciones_tipo_hogar'] = lista_coordinaciones_tipo_hogar
                    data['tipos_hogares'] = TipoHogar.objects.all()
                    return render(request, "estadisticas/graf_tipohogar.html", data)
                except Exception as ex:
                    pass

            elif action == 'graph_nivelescolar_jefehogar':
                try:
                    # CONFORMACION DE GRAFICOS PARA MOSTRAR EL NIVEL ESCOLAR DE LOS JEFES DE HOGAR
                    data['title'] = u'Niveles de Escolaridad Jefes de Hogar'
                    coordinaciones = Coordinacion.objects.all().order_by('id')
                    carreras = Carrera.objects.all()
                    lista_carreras_nivel_estudios = []
                    for c in carreras:
                        lista_nivelestudio = []
                        for nivel in NivelEstudio.objects.all():
                            lista_nivelestudio.append(cantidad_nivel_educacion_jefehogar_carrera(nivel, c))
                        lista_carreras_nivel_estudios.append((c.alias, lista_nivelestudio))
                    data['carreras'] = carreras
                    data['lista_carreras_nivel_estudios'] = lista_carreras_nivel_estudios
                    lista_coordinaciones_nivel_estudios = []
                    for c in coordinaciones:
                        lista_nivelestudio = []
                        for nivel in NivelEstudio.objects.all():
                            lista_nivelestudio.append(cantidad_nivel_educacion_jefehogar_coordinacion(nivel, c))
                        lista_coordinaciones_nivel_estudios.append((c.nombre, lista_nivelestudio))
                    data['coordinaciones'] = coordinaciones
                    data['lista_coordinaciones_nivel_estudios'] = lista_coordinaciones_nivel_estudios
                    data['niveles_estudios'] = NivelEstudio.objects.all()
                    return render(request, "estadisticas/graf_nivelescolar_jefeh.html", data)
                except Exception as ex:
                    pass

            elif action == 'graph_dependencia_carrera':
                try:
                    # CONFORMACION DE GRAFICOS DE DEPENDENCIAS ECONOMICAS POR CARRERAS
                    data['title'] = u'Dependencia Economica por carreras'
                    carreras = Carrera.objects.all()
                    lista_carreras_dependencia = []
                    for c in carreras:
                        lista_carreras_dependencia.append((c.alias, cantidad_sidependientes_carrera(c), cantidad_nodependientes_carrera(c), c.id))
                    data['lista_carreras_dependencia'] = lista_carreras_dependencia
                    return render(request, "estadisticas/graf_depend_carrera.html", data)
                except Exception as ex:
                    pass

            elif action == 'graph_dependencia_coordinacion':
                try:
                    # CONFORMACION DE GRAFICOS DE DEPENDENCIAS ECONOMICAS POR COORDINACIONES
                    data['title'] = u'Dependencia Economica por coordinaciones'
                    coordinaciones = Coordinacion.objects.all().order_by('id')
                    lista_coordinaciones_dependencia = []
                    for c in coordinaciones:
                        lista_coordinaciones_dependencia.append((c.nombre, cantidad_sidependientes_coordinacion(c), cantidad_nodependientes_coordinacion(c), c.id))
                    data['lista_coordinaciones_dependencia'] = lista_coordinaciones_dependencia
                    return render(request, "estadisticas/graf_depend_coord.html", data)
                except Exception as ex:
                    pass

            elif action == 'graph_cabezaf_carrera':
                try:
                    # CONFORMACION DE GRAFICOS DE CABEZAS DE FAMILIA POR CARRERAS
                    data['title'] = u'Estudiantes cabezas de familias por carreras'
                    carreras = Carrera.objects.all()
                    lista_carreras_cabezasf = []
                    for c in carreras:
                        lista_carreras_cabezasf.append((c.alias, cantidad_sicabezasf_carrera(c), cantidad_nocabezasf_carrera(c), c.id))
                    data['lista_carreras_cabezasf'] = lista_carreras_cabezasf
                    return render(request, "estadisticas/graf_cabezas_carrera.html", data)
                except Exception as ex:
                    pass

            elif action == 'graph_cabezaf_coordinacion':
                try:
                    # CONFORMACION DE GRAFICOS DE CABEZAS DE FAMILIA POR COORDINACIONES
                    data['title'] = u'Estudiantes cabezas de familias por coordinaciones'
                    coordinaciones = Coordinacion.objects.all().order_by('id')
                    lista_coordinaciones_cabezasf = []
                    for c in coordinaciones:
                        lista_coordinaciones_cabezasf.append((c.nombre, cantidad_sicabezasf_coordinacion(c), cantidad_nocabezasf_coordinacion(c), c.id))
                    data['lista_coordinaciones_cabezasf'] = lista_coordinaciones_cabezasf
                    return render(request, 'estadisticas/graf_cabezas_coord.html', data)
                except Exception as ex:
                    pass

            elif action == 'tablasegmentodia':
                try:
                    data['title'] = u'Tablas de operaciones del dia'
                    hoy = datetime.now().today()
                    data['hoy'] = hoy
                    data['total_efectivo_dia'] = total_efectivo_dia(hoy)
                    data['cantidad_cheques_dia'] = cantidad_cheques_dia(hoy)
                    data['total_cheque_dia'] = total_cheque_dia(hoy)
                    data['cantidad_tarjetas_dia'] = cantidad_tarjetas_dia(hoy)
                    data['total_tarjeta_dia'] = total_tarjeta_dia(hoy)
                    data['cantidad_depositos_dia'] = cantidad_depositos_dia(hoy)
                    data['total_deposito_dia'] = total_deposito_dia(hoy)
                    data['cantidad_transferencias_dia'] = cantidad_transferencias_dia(hoy)
                    data['total_transferencia_dia'] = total_transferencia_dia(hoy)
                    data['cantidad_facturas_dia'] = cantidad_facturas_dia(hoy)
                    data['total_dia'] = total_dia(hoy)
                    data['sesiones'] = [x for x in SesionCaja.objects.filter(fecha=hoy).order_by('caja__nombre') if x.total_sesion()]
                    data['carreras'] = Carrera.objects.filter(activa=True)
                    return render(request, "estadisticas/tablasegmento.html", data)
                except Exception as ex:
                    pass

            elif action == 'tablasegmentomes':
                try:
                    data['title'] = u'Tablas de operaciones del mes'
                    hoy = datetime.now().date()
                    ultimodia = calendar.monthrange(hoy.year, hoy.month)[1]
                    data['hoy'] = hoy
                    data['total_efectivo_mes'] = total_efectivo_mes()
                    lista_ingresos_mes = []
                    for x in range(1, ultimodia + 1):
                        fecha = date(hoy.year, hoy.month, x)
                        a = [date(hoy.year, hoy.month, x),
                             total_efectivo_dia(fecha),
                             cantidad_cheques_dia(fecha),
                             total_cheque_dia(fecha),
                             cantidad_tarjetas_dia(fecha),
                             total_tarjeta_dia(fecha),
                             cantidad_depositos_dia(fecha),
                             total_deposito_dia(fecha),
                             cantidad_transferencias_dia(fecha),
                             total_transferencia_dia(fecha),
                             cantidad_facturas_dia(fecha),
                             total_dia(fecha)]
                        lista_ingresos_mes.append(a)
                    data['lista_ingresos_mes'] = lista_ingresos_mes
                    inicio = date(hoy.year, hoy.month, 1)
                    fin = date(hoy.year, hoy.month, ultimodia)
                    data['total_efectivo_mes'] = total_efectivo_rango(inicio, fin)
                    data['cantidad_cheques_mes'] = cantidad_cheques_rango(inicio, fin)
                    data['total_cheque_mes'] = total_cheque_rango(inicio, fin)
                    data['cantidad_tarjetas_mes'] = cantidad_tarjetas_rango(inicio, fin)
                    data['total_tarjeta_mes'] = total_tarjeta_rango(inicio, fin)
                    data['cantidad_depositos_mes'] = cantidad_depositos_rango(inicio, fin)
                    data['total_deposito_mes'] = total_deposito_rango(inicio, fin)
                    data['cantidad_transferencias_mes'] = cantidad_transferencias_rango(inicio, fin)
                    data['total_transferencia_mes'] = total_transferencia_rango(inicio, fin)
                    data['cantidad_facturas_mes'] = cantidad_facturas_rango(inicio, fin)
                    data['total_mes'] = total_rango(inicio, fin)
                    lista_ingresos_carrera_mes = []
                    for carrera in Carrera.objects.filter(activa=True):
                        lista_ingresos_carrera_mes.append([carrera, carrera.cantidad_facturas_rango_fechas(inicio, fin), carrera.total_pagos_rango(inicio, fin)])
                    data['lista_ingresos_carrera_mes'] = lista_ingresos_carrera_mes
                    return render(request, "estadisticas/tablasegmentomes.html", data)
                except Exception as ex:
                    pass

            elif action == 'tablasegmentorango':
                try:
                    data['title'] = u'Tablas de operaciones de un rango de fecha'
                    inicio = convertir_fecha_invertida(request.GET['inicio'])
                    fin = convertir_fecha_invertida(request.GET['fin'])
                    data['total_efectivo_rango'] = total_efectivo_rango(inicio, fin)
                    lista_ingresos_rango = []
                    fecha = inicio
                    for x in range(1, (fin - inicio).days + 2):
                        a = [fecha,
                             total_efectivo_dia(fecha),
                             cantidad_cheques_dia(fecha),
                             total_cheque_dia(fecha),
                             cantidad_tarjetas_dia(fecha),
                             total_tarjeta_dia(fecha),
                             cantidad_depositos_dia(fecha),
                             total_deposito_dia(fecha),
                             cantidad_transferencias_dia(fecha),
                             total_transferencia_dia(fecha),
                             cantidad_facturas_dia(fecha),
                             total_dia(fecha)]
                        lista_ingresos_rango.append(a)
                        fecha = (datetime(inicio.year, inicio.month, inicio.day, 0, 0, 0) + timedelta(days=x)).date()
                    data['lista_ingresos_rango'] = lista_ingresos_rango
                    data['total_efectivo_rango'] = total_efectivo_rango(inicio, fin)
                    data['cantidad_cheques_rango'] = cantidad_cheques_rango(inicio, fin)
                    data['total_cheque_rango'] = total_cheque_rango(inicio, fin)
                    data['cantidad_tarjetas_rango'] = cantidad_tarjetas_rango(inicio, fin)
                    data['total_tarjeta_rango'] = total_tarjeta_rango(inicio, fin)
                    data['cantidad_depositos_rango'] = cantidad_depositos_rango(inicio, fin)
                    data['total_deposito_rango'] = total_deposito_rango(inicio, fin)
                    data['cantidad_transferencias_rango'] = cantidad_transferencias_rango(inicio, fin)
                    data['total_transferencia_rango'] = total_transferencia_rango(inicio, fin)
                    data['cantidad_facturas_rango'] = cantidad_facturas_rango(inicio, fin)
                    data['total_rango'] = total_rango(inicio, fin)
                    lista_ingresos_carrera_rango = []
                    for carrera in Carrera.objects.filter(activa=True):
                        lista_ingresos_carrera_rango.append([carrera, carrera.cantidad_facturas_rango_fechas(inicio, fin), carrera.total_pagos_rango(inicio, fin)])
                    data['lista_ingresos_carrera_rango'] = lista_ingresos_carrera_rango
                    return render(request, "estadisticas/tablasegmentorango.html", data)
                except Exception as ex:
                    pass

            elif action == 'tablasegmentodeudacoordvenc':
                try:
                    data['title'] = u'Tablas de Deudas por Coordinaciones'
                    data['coordinaciones'] = Coordinacion.objects.all()
                    data['valor_total_deudores_activos_30dias'] = valor_total_deudores_activos_30dias()
                    data['valor_total_deudores_activos_31_90dias'] = valor_total_deudores_activos_31_90dias()
                    data['valor_total_deudores_activos_mas_90dias'] = valor_total_deudores_activos_mas_90dias()
                    data['cantidad_total_deudores'] = cantidad_total_deudores()
                    data['valor_deudores_activos_total'] = valor_deudores_activos_total()
                    return render(request, "estadisticas/tablasegmentodeudacoordvenc.html", data)
                except Exception as ex:
                    pass



            elif action == 'tablasegmentodeudacarvenc':
                try:
                    data['title'] = u'Tablas de Deudas por Carreras'
                    data['carreras'] = Carrera.objects.all()
                    data['valor_total_deudores_activos_30dias'] = valor_total_deudores_activos_30dias()
                    data['valor_total_deudores_activos_31_90dias'] = valor_total_deudores_activos_31_90dias()
                    data['valor_total_deudores_activos_mas_90dias'] = valor_total_deudores_activos_mas_90dias()
                    data['cantidad_total_deudores'] = cantidad_total_deudores()
                    data['valor_deudores_activos_total'] = valor_deudores_activos_total()
                    return render(request, "estadisticas/tablasegmentodeudacarvenc.html", data)
                except Exception as ex:
                    pass

            elif action == 'tablasegmentodeudacarfut':
                try:
                    data['title'] = u'Valores a pagar por Carreras'
                    data['carreras'] = Carrera.objects.all()
                    data['valor_total_apagar_activos_30dias'] = valor_total_apagar_activos_30dias()
                    data['valor_total_apagar_activos_31_90dias'] = valor_total_apagar_activos_31_90dias()
                    data['valor_total_apagar_activos_mas_90dias'] = valor_total_apagar_activos_mas_90dias()
                    data['cantidad_total_apagar'] = cantidad_total_apagar()
                    data['valor_apagar_activos_total'] = valor_apagar_activos_total()
                    return render(request, "estadisticas/tablasegmentodeudacarfut.html", data)
                except Exception as ex:
                    pass

            elif action == 'tablasegmentodeudacoorfut':
                try:
                    data['title'] = u'Valores a pagar por Coordinaciones'
                    data['coordinaciones'] = Coordinacion.objects.all()
                    data['valor_total_apagar_activos_30dias'] = valor_total_apagar_activos_30dias()
                    data['valor_total_apagar_activos_31_90dias'] = valor_total_apagar_activos_31_90dias()
                    data['valor_total_apagar_activos_mas_90dias'] = valor_total_apagar_activos_mas_90dias()
                    data['cantidad_total_apagar'] = cantidad_total_apagar()
                    data['valor_apagar_activos_total'] = valor_apagar_activos_total()
                    return render(request, "estadisticas/tablasegmentodeudacoorfut.html", data)
                except Exception as ex:
                    pass

            elif action == 'tablasegmentodeudacartotal':
                try:
                    data['title'] = u'Tabla de Totales Valores Vencidos y a Pagar por Carreras'
                    data['carreras'] = Carrera.objects.all()
                    data['valor_apagar_activos_total'] = valor_apagar_activos_total()
                    data['valor_deudores_activos_total'] = valor_deudores_activos_total()
                    data['valor_deudas_activos_total'] = valor_deudas_activos_total()
                    return render(request, "estadisticas/tablasegmentodeudacartotal.html", data)
                except Exception as ex:
                    pass

            elif action == 'tablasegmentodeudacoortotal':
                try:
                    data['title'] = u'Tabla de Totales Valores Vencidos y a Pagar por Carreras'
                    data['coordinaciones'] = Coordinacion.objects.all()
                    data['valor_apagar_activos_total'] = valor_apagar_activos_total()
                    data['valor_deudores_activos_total'] = valor_deudores_activos_total()
                    data['valor_deudas_activos_total'] = valor_deudas_activos_total()
                    return render(request, "estadisticas/tablasegmentodeudacoortotal.html", data)
                except Exception as ex:
                    pass

            elif action == 'segmentoacademicocarr':
                try:
                    data['title'] = u'Matriculados por carreras segun genero'
                    data['carreras'] = [c for c in Carrera.objects.filter(status=True).order_by('id') if c.cantidad_matriculados(periodo)]
                    data['total_matriculados_mujeres'] = total_matriculados_mujeres(periodo)
                    data['total_matriculados_hombres'] = total_matriculados_hombres(periodo)
                    data['total_matriculados'] = total_matriculados(periodo)
                    return render(request, "estadisticas/tablasegmentoacademicocarr.html", data)
                except Exception as ex:
                    pass

            elif action == 'matriculadosfacultad':
                try:
                    data['title'] = u'Matriculados por facultad'

                    data['coordinaciones'] = coordina = Coordinacion.objects.filter(status=True, id__in=[1,2,3,4,5]).order_by('id')
                    data['total_matriculados'] = sum([c.cantidad_matriculados_periodo(periodo) for c in coordina])

                    return render(request, "estadisticas/tablamatriculadosfacul.html", data)
                except Exception as ex:
                    pass



            elif action == 'reporte_matriculadosfacultad':
                try:
                    nombreArchivo = 'MATRICULADOS DE ' + periodo.nombre.__str__() + '.xlsx'

                    libro = xlsxwriter.Workbook(nombreArchivo)
                    ws = libro.add_worksheet()
                    formatoTitulo = libro.add_format({
                        'bold': True,
                        'border': 1,
                        'align': 'center',
                        'valign': 'vcenter',
                    })
                    formatosubtitulo = libro.add_format({
                        'bold': True,
                        'valign': 'vcenter',
                    })

                    ws.merge_range('B1:J1', 'UNIVERSIDAD ESTATAL DE MILAGRO', formatoTitulo)

                    ws.write(4, 0, 'N.',formatosubtitulo)
                    ws.write(4, 1, 'PERIODO',formatosubtitulo)
                    ws.write(4, 2, 'NIVEL_CRE', formatosubtitulo)
                    ws.write(4, 3, 'NIVEL_MAT', formatosubtitulo)
                    ws.write(4, 4, 'SECCION', formatosubtitulo)
                    ws.write(4, 5, 'CEDULA', formatosubtitulo)
                    ws.write(4, 6, 'APELLIDOS', formatosubtitulo)
                    ws.write(4, 7, 'NOMBRES', formatosubtitulo)
                    ws.write(4, 8, 'SEXO', formatosubtitulo)
                    ws.write(4, 9, 'FECHANACIMIENTO', formatosubtitulo)
                    ws.write(4, 10, 'EMAIL', formatosubtitulo)
                    ws.write(4, 11, 'EMAILINST', formatosubtitulo)
                    ws.write(4, 12, 'COORDINACION', formatosubtitulo)
                    ws.write(4, 13, 'CARRERA', formatosubtitulo)
                    ws.write(4, 14, 'COD. SENESCYT', formatosubtitulo)
                    ws.write(4, 15, 'TELEFONO', formatosubtitulo)
                    ws.write(4, 16, 'INSCRIPCION', formatosubtitulo)
                    ws.write(4, 17, '# MATRICULAS', formatosubtitulo)
                    ws.write(4, 18, 'LGTBI', formatosubtitulo)
                    ws.write(4, 19, 'ETNIA', formatosubtitulo)
                    ws.write(4, 20, 'NACIONALIDAD', formatosubtitulo)
                    ws.write(4, 21, 'PAIS', formatosubtitulo)
                    ws.write(4, 22, 'PROVINCIA', formatosubtitulo)
                    ws.write(4, 23, 'CANTON', formatosubtitulo)
                    ws.write(4, 24, 'DIRECCION', formatosubtitulo)
                    ws.write(4, 25, 'ESTADO SOCIO ECONOMICO', formatosubtitulo)
                    ws.write(4, 26, 'REALIZADO', formatosubtitulo)
                    ws.write(4, 27, 'FECHA INICIO PRIMER NIVEL', formatosubtitulo)
                    ws.write(4, 28, 'FECHA CONVALIDACION', formatosubtitulo)
                    ws.write(4, 29, 'TIENE DISCAPACIDAD', formatosubtitulo)
                    ws.write(4, 30, 'DISCAPACIDAD', formatosubtitulo)
                    ws.write(4, 31, 'ID MATRICULA', formatosubtitulo)
                    ws.write(4, 32, 'TIPO MATRICULA', formatosubtitulo)
                    ws.write(4, 33, 'TIPO ESTUDIANTE', formatosubtitulo)
                    ws.write(4, 34, 'CONTACTO EMERGENCIA', formatosubtitulo)
                    ws.write(4, 35, 'COLEGIO', formatosubtitulo)
                    ws.write(4, 36, 'MODALIDAD', formatosubtitulo)
                    ws.write(4, 37, 'PPL', formatosubtitulo)
                    ws.write(4, 38, 'ACEPTA MATRÍCULA', formatosubtitulo)

                    ws.set_column(0, 0, 5)
                    ws.set_column(1, 1, 30)
                    ws.set_column(2, 2, 10)
                    ws.set_column(3, 3, 10)
                    ws.set_column(4, 9, 25)
                    ws.set_column(10, 10, 30)
                    ws.set_column(11, 11, 30)
                    ws.set_column(12, 38, 20)

                    periodoid = str(request.session['periodo'].id)
                    cursor = connections['sga_select'].cursor()

                    listaestudiante = """
                                    select peri.nombre as PERIODO, nimalla.nombre as NIVEL_CRE,
                                    sesion.nombre as SECCION,
                                    perso.cedula, perso.apellido1 || ' ' || perso.apellido2 as apellidos, 
                                    perso.nombres as nompersona, se.nombre as sexo, 
                                    perso.nacimiento as fechanacimiento,perso.email,perso.emailinst, coor.nombre as facultad, carr.nombre as carrera, 
                                    (select ma.codigo from sga_inscripcionmalla insmalla inner join sga_malla ma on insmalla.malla_id=ma.id where insmalla.inscripcion_id=ins.id and insmalla.status=True), 
                                    perso.telefono,ins.id as INSCRIPCION, 
                                    CASE WHEN perso.lgtbi = True THEN 'SI' else 'NO' END as lgtbi, 
                                    (select rasa.nombre from sga_perfilinscripcion perfil inner join sga_raza rasa on perfil.raza_id=rasa.id where perfil.persona_id=perso.id and perfil.status=True) as etnia, 
                                    perso.nacionalidad, pais.nombre as pais, provincia.nombre as provincia, canton.nombre as canton, 
                                    perso.direccion || ' ' || perso.direccion2 as direccion, 
                                    (select gru.codigo || ' ' || gru.nombre from socioecon_fichasocioeconomicainec fi inner join socioecon_gruposocioeconomico gru on fi.grupoeconomico_id=gru.id where fi.persona_id=perso.id and fi.status=True) as gruposocioeconomico, 
                                    ins.fechainicioprimernivel,ins.fechainicioconvalidacion, 
                                    (select CASE WHEN perfilin.tienediscapacidad = True THEN 'SI' else 'NO' END  from sga_perfilinscripcion perfilin where perfilin.persona_id=perso.id and perfilin.status=True) as tienediscapacidad, 
                                    (select disca.nombre  from sga_perfilinscripcion perfilin,sga_discapacidad disca where perfilin.persona_id=perso.id and perfilin.tipodiscapacidad_id=disca.id and perfilin.status=True) as discapacidad, 
                                    matri.id as idmatricula, CASE WHEN gruso.tipomatricula=1 THEN 'REGULAR' WHEN gruso.tipomatricula=2 THEN 'IRREGULAR' END as tipomatricula, 
                                    tima.nombre as tipoestudiante, ext.contactoemergencia as contactoemergencia, 
                                    uni.nombre as unidadeducativa, modal.nombre as modalidad,
                                    (select count(*) from sga_matricula matr where matr.inscripcion_id=ins.id and matr.status=True
                                    and matr.retiradomatricula=False) as veces_matricula , perso.ppl, matri.termino
                                    from sga_matricula matri                                   
                                     
                                    
                                    inner join sga_tipomatricula tima on tima.id=matri.tipomatricula_id and tima.status=True
                                    
                                    inner join sga_nivel ni on matri.nivel_id=ni.id and ni.status=True
                                    
                                    
                                    INNER JOIN sga_nivellibrecoordinacion  ON ni.id = sga_nivellibrecoordinacion.nivel_id                                  
                                    
                                    inner join sga_inscripcion ins on matri.inscripcion_id=ins.id and ins.status=True
                                    left join sga_coordinacion coor on coor.id=ins.coordinacion_id                                 
                                    
                                    left join sga_carrera carr on carr.id=ins.carrera_id
                                    left join sga_institucionescolegio uni on uni.id=ins.unidadeducativa_id and uni.status=True
                                    left join sga_matriculagruposocioeconomico gruso on gruso.matricula_id=matri.id and gruso.status=True
                                    inner join sga_persona perso on ins.persona_id=perso.id
                                    left join med_personaextension ext on ext.persona_id=perso.id and ext.status=True
                                    left join sga_sexo se on se.id=perso.sexo_id and se.status=True
                                    left join sga_pais pais on perso.pais_id=pais.id and pais.status=True
                                    left join sga_provincia provincia on perso.provincia_id=provincia.id and provincia.status=True
                                    left join sga_canton canton on perso.canton_id=canton.id and canton.status=True
                                    inner join sga_periodo peri on ni.periodo_id=peri.id
                                    inner join sga_nivelmalla nimalla on matri.nivelmalla_id=nimalla.id 
                                    left join sga_sesion sesion on ins.sesion_id=sesion.id
                                    inner join sga_modalidad modal on ins.modalidad_id=modal.id
                                    where  matri.status=True and matri.retiradomatricula=False
                                    and ni.periodo_id= %s
												AND sga_nivellibrecoordinacion.coordinacion_id IN (1,2,3,4,5)
												and matri.estado_matricula IN (2, 3)"""% (periodoid)

                    cursor.execute(listaestudiante)
                    results = cursor.fetchall()
                    a = 4

                    for per in results:
                        a += 1
                        ws.write(a, 0, a - 4)
                        ws.write(a, 1, '%s' % per[0])
                        ws.write(a, 2, '%s' % per[1])
                        ws.write(a, 3, '%s' % per[1])
                        ws.write(a, 4, '%s' % per[2])
                        ws.write(a, 5, '%s' % per[3])
                        ws.write(a, 6, '%s' % per[4])
                        ws.write(a, 7, '%s' % per[5])
                        ws.write(a, 8, '%s' % per[6])
                        # ws.write(a, 9, per[7], date_format)
                        ws.write(a, 10, '%s' % per[8])
                        ws.write(a, 11, '%s' % per[9])
                        ws.write(a, 12, '%s' % per[10])
                        ws.write(a, 13, '%s' % per[11])
                        ws.write(a, 14, '%s' % per[12])
                        ws.write(a, 15, '%s' % per[13])
                        ws.write(a, 16, '%s' % per[14])
                        ws.write(a, 17, '%s' % per[33])
                        ws.write(a, 18, '%s' % per[15])
                        ws.write(a, 19, '%s' % per[16])
                        ws.write(a, 20, '%s' % per[17])
                        ws.write(a, 21, '%s' % per[18])
                        ws.write(a, 22, '%s' % per[19])
                        ws.write(a, 23, '%s' % per[20])
                        ws.write(a, 24, '%s' % per[21])
                        ws.write(a, 25, '%s' % per[22])
                        ws.write(a, 26, 'SI')
                        ws.write(a, 27, '%s' % per[23])
                        ws.write(a, 28, '%s' % per[24])
                        ws.write(a, 29, '%s' % per[25])
                        ws.write(a, 30, '%s' % per[26])
                        ws.write(a, 31, '%s' % per[27])
                        ws.write(a, 32, '%s' % per[28])
                        ws.write(a, 33, '%s' % per[29])
                        ws.write(a, 34, '%s' % per[30])
                        ws.write(a, 35, '%s' % per[31])
                        ws.write(a, 36, '%s' % per[32])
                        ws.write(a, 37, '%s' % "SI" if per[34] else "NO")
                        ws.write(a, 38, '%s' % "SI" if per[35] else "NO")


                    libro.close()
                    # Crear respuesta HttpResponse
                    response = HttpResponse(content_type="application/ms-excel")
                    content = "attachment; filename = {0}".format(nombreArchivo)
                    response['Content-Disposition'] = content
                    response.write(open(nombreArchivo, 'rb').read())

                    return response

                except Exception as ex:
                    pass


            elif action == 'segmentoacademicocoor':
                try:
                    data['title'] = u'Matriculados por coordinaciones segun genero'
                    data['coordinaciones'] = [c for c in Coordinacion.objects.all().order_by('id') if c.cantidad_matriculados(periodo)]
                    data['total_matriculados_mujeres'] = total_matriculados_mujeres(periodo)
                    data['total_matriculados_hombres'] = total_matriculados_hombres(periodo)
                    data['total_matriculados'] = total_matriculados(periodo)
                    return render(request, "estadisticas/tablasegmentoacademicocoor.html", data)
                except Exception as ex:
                    pass

            elif action == 'matriculadosrangoedadescarr':
                try:
                    data['title'] = u'Matriculados por carreras segun rango de edades'
                    data['carreras'] = [c for c in Carrera.objects.filter(status=True).order_by('id') if c.cantidad_matriculados(periodo)]
                    data['matriculados_menor_30'] = total_matriculados_menor_30(periodo)
                    data['matriculados_31_40'] = total_matriculados_31_40(periodo)
                    data['matriculados_41_50'] = total_matriculados_41_50(periodo)
                    data['matriculados_51_60'] = total_matriculados_51_60(periodo)
                    data['matriculados_mayor_61'] = total_matriculados_mayor_61(periodo)
                    data['total_matriculados'] = total_matriculados(periodo)
                    return render(request, "estadisticas/matriculadosrangoedadescarr.html", data)
                except Exception as ex:
                    pass

            elif action == 'mapeoventascomercializacion':
                try:
                    data['title'] = u'Mapeo de ventas por Provincia'
                    lista_provincias = []
                    lista_cantones = []
                    cantones_con_ventas = []

                    # listado_pagos = []
                    ventas_prov = []
                    listado_reportadas = []
                    ventas_prov_nf = []
                    ventas_canto_nf = []

                    maestria = cohorte = asesorcomercial = pais = provincia = canton = anio = 0
                    desde = hasta = ''

                    filtro = Q(status=True)
                    if 'maestria' in request.GET:
                        maestria = int(request.GET['maestria'])
                    if 'cohorte' in request.GET:
                        cohorte = int(request.GET['cohorte'])
                    if 'asesor' in request.GET:
                        asesorcomercial = int(request.GET['asesor'])
                    if 'pais' in request.GET:
                        pais = int(request.GET['pais'])
                    if 'provincia' in request.GET:
                        provincia = int(request.GET['provincia'])
                    if 'canton' in request.GET:
                        canton = int(request.GET['canton'])
                    if 'desde' in request.GET:
                        desde = request.GET['desde']
                    if 'hasta' in request.GET:
                        hasta = request.GET['hasta']

                    url_vars = ' '
                    cohortes = provincias = cantones= None
                    lis_canto = lis_cancan = []
                    tot = 0

                    if maestria > 0:
                        data['maestria'] = maestria
                        filtro = filtro & Q(inscripcioncohorte__cohortes__maestriaadmision__id=maestria)
                        cohortes = CohorteMaestria.objects.filter(maestriaadmision__id=maestria)
                        url_vars += "&maestria={}".format(maestria)

                    if cohorte > 0:
                        data['cohorte'] = cohorte
                        filtro = filtro & Q(inscripcioncohorte__cohortes__id=cohorte)
                        url_vars += "&cohorte={}".format(cohorte)

                    if asesorcomercial > 0:
                        data['asesor'] = asesorcomercial
                        filtro = filtro & Q(asesor=asesorcomercial)
                        url_vars += "&asesor={}".format(asesorcomercial)

                    if pais > 0:
                        data['pais'] = pais
                        filtro = filtro & Q(inscripcioncohorte__inscripcionaspirante__persona__pais__id=pais)
                        provincias = Provincia.objects.filter(pais__id=pais)
                        url_vars += "&pais={}".format(pais)

                    if provincia > 0:
                        data['provincia'] = provincia
                        filtro = filtro & Q(inscripcioncohorte__inscripcionaspirante__persona__provincia__id=provincia)
                        cantones = Canton.objects.filter(provincia__id=provincia)
                        url_vars += "&provincia={}".format(provincia)

                    if desde and hasta:
                        data['desde'] = desde
                        data['hasta'] = hasta
                        url_vars += "&desde={}".format(desde)
                        url_vars += "&hasta={}".format(hasta)
                        filtro = filtro & Q(fecha__range=(desde, hasta))

                    elif desde:
                        data['desde'] = desde
                        url_vars += "&desde={}".format(desde)
                        filtro = filtro & Q(fecha__gte=desde)
                    elif hasta:
                        data['hasta'] = hasta
                        url_vars += "&hasta={}".format(hasta)
                        filtro = filtro & Q(fecha__lte=hasta)

                    if maestria == 0 and cohorte == 0 and asesorcomercial == 0 and pais == 0 and provincia == 0 and canton == 0 and desde == '' and hasta == '':
                        leads = []
                        provincias_con_ventas = []
                    else:
                        leads = VentasProgramaMaestria.objects.filter(filtro).exclude(valida=False).order_by('-id')
                        listado_reportadas.append(leads.values_list('inscripcioncohorte__id', flat=True))
                        tot = leads.count()
                        provincias_con_ventas = leads.values_list('inscripcioncohorte__inscripcionaspirante__persona__provincia__id', flat=True).order_by('inscripcioncohorte__inscripcionaspirante__persona__provincia__nombre').distinct().exclude(inscripcioncohorte__inscripcionaspirante__persona__provincia__isnull=True)
                        cantones_con_ventas = leads.values_list('inscripcioncohorte__inscripcionaspirante__persona__canton__id', flat=True).order_by('inscripcioncohorte__inscripcionaspirante__persona__canton__nombre').distinct().exclude(inscripcioncohorte__inscripcionaspirante__persona__provincia__isnull=True)

                    dicc = {}
                    for province in provincias_con_ventas:
                        provi = Provincia.objects.get(status=True, pk=province)

                        qty2 = VentasProgramaMaestria.objects.filter(inscripcioncohorte__id__in=listado_reportadas, inscripcioncohorte__inscripcionaspirante__persona__provincia=provi).exclude(valida=False).count()
                        # ventas_prov_nf.append(qty2)

                        if provi.id == 25:
                            # lista_provincias.append('STD. DE LOS TSACHILAS')
                            dicc = {'provincia': 'STD. DE LOS TSACHILAS', 'cantidad': qty2}
                            lista_provincias.append(dicc)
                        else:
                            dicc = {'provincia': provi.nombre, 'cantidad': qty2}
                            lista_provincias.append(dicc)
                            # lista_provincias.append(provi.nombre)

                    if len(listado_reportadas) > 0 and provincia == 0:
                        # lista_provincias.append('SIN PROVINCIA')
                        qtnulo = VentasProgramaMaestria.objects.filter(inscripcioncohorte__id__in=listado_reportadas, inscripcioncohorte__inscripcionaspirante__persona__provincia__isnull=True).exclude(valida=False).count()
                        # ventas_prov_nf.append(qtnulo)
                        dicc = {'provincia': 'SIN PROVINCIA', 'cantidad': qtnulo}
                        lista_provincias.append(dicc)
                        data['noprov'] = Provincia.objects.get(status=True, pk=1)

                    newlist = sorted(lista_provincias, key=lambda d: d['cantidad'], reverse=True)
                    lis_provi = [d['provincia'] for d in newlist]
                    lis_can = [d['cantidad'] for d in newlist]

                    if provincia > 0:
                        dicc2 = {}
                        for cantone in cantones_con_ventas:
                            canto = Canton.objects.get(status=True, pk=cantone)

                            qty3 = VentasProgramaMaestria.objects.filter(inscripcioncohorte__id__in=listado_reportadas, inscripcioncohorte__inscripcionaspirante__persona__canton=canto).exclude(valida=False).count()
                            # ventas_canto_nf.append(qty3)
                            dicc2 = {'canton': canto.nombre, 'cantidad': qty3}
                            lista_cantones.append(dicc2)

                        # lista_cantones.append('SIN CANTÓN')
                        qtnulo = VentasProgramaMaestria.objects.filter(inscripcioncohorte__id__in=listado_reportadas, inscripcioncohorte__inscripcionaspirante__persona__canton__isnull=True).exclude(valida=False).count()
                        dicc2 = {'canton': 'SIN CANTÓN', 'cantidad': qtnulo}
                        lista_cantones.append(dicc2)
                        # ventas_canto_nf.append(qtnulo)
                        data['notown'] = Canton.objects.get(status=True, pk=1)

                        newlist = sorted(lista_cantones, key=lambda d: d['cantidad'], reverse=True)
                        lis_canto = [d['canton'] for d in newlist]
                        lis_cancan = [d['cantidad'] for d in newlist]

                    # lista_provincias.append('SIN PROVINCIA')
                    data['listado_reportadas'] = listado_reportadas
                    data['provincias'] = lis_provi
                    data['cantons'] = lis_canto
                    data['provinces'] = Provincia.objects.filter(status=True, id__in=provincias_con_ventas).order_by('nombre')
                    data['towns'] = Canton.objects.filter(status=True, id__in=cantones_con_ventas).order_by('nombre')
                    data['cant_ventas_nf'] = lis_can
                    data['cant_ventas_canto'] = lis_cancan
                    # data['total_ventas'] = InscripcionCohorte.objects.filter(id__in=listado_reportadas).count()
                    data['maestrialist'] = MaestriasAdmision.objects.filter(status=True, carrera__coordinacion__id=7)
                    data['cohorteslist'] = cohortes
                    data['asesorlist'] = AsesorComercial.objects.filter(status=True)
                    data['paislist'] = Pais.objects.filter(status=True)
                    data['provincialist'] = provincias
                    data['cantonlist'] = cantones
                    data['url_vars'] = url_vars

                    data['total_ventas_menor_30'] = total_ventas_edades(listado_reportadas, 'menor', 30)
                    data['total_ventas_31_40'] = total_ventas_edades(listado_reportadas, 'doble', 31, 40)
                    data['total_ventas_41_50'] = total_ventas_edades(listado_reportadas, 'doble', 41, 50)
                    data['total_ventas_51_60'] = total_ventas_edades(listado_reportadas, 'doble', 51, 60)
                    data['total_ventas_mayor_61'] = total_ventas_edades(listado_reportadas, 'mayor', 61)

                    data['total_ventas_prov_mujeres'] = total_ventas_prov_sexo(listado_reportadas, 'femenino')
                    data['total_ventas_prov_hombres'] = total_ventas_prov_sexo(listado_reportadas, 'masculino')

                    data['total_ventas'] = tot
                    return render(request, "estadisticas/mapeoventascomercializacion.html", data)
                except Exception as ex:
                    pass

            elif action == 'mapeoventasasesor':
                try:
                    data['title'] = u'Mapeo de ventas por Asesor Comercial'
                    listado_reportadas = []
                    # ventas_ase = []
                    listado_asesor = []
                    tot = 0

                    maestria = cohorte = asesorcomercial = pais = provincia = canton = anio = rol = 0
                    desde = hasta = ''

                    filtro = Q(status=True)
                    if 'maestria' in request.GET:
                        maestria = int(request.GET['maestria'])
                    if 'cohorte' in request.GET:
                        cohorte = int(request.GET['cohorte'])
                    if 'asesor' in request.GET:
                        asesorcomercial = int(request.GET['asesor'])
                    if 'pais' in request.GET:
                        pais = int(request.GET['pais'])
                    if 'provincia' in request.GET:
                        provincia = int(request.GET['provincia'])
                    if 'canton' in request.GET:
                        canton = int(request.GET['canton'])
                    if 'desde' in request.GET:
                        desde = request.GET['desde']
                    if 'hasta' in request.GET:
                        hasta = request.GET['hasta']
                    if 'rol' in request.GET:
                        rol = int(request.GET['rol'])

                    url_vars = ' '
                    cohortes = provincias = cantones= None
                    primero = segundo = tercero = ''
                    prican = segcan = tercan = 0
                    first = second = third = None

                    if maestria > 0:
                        data['maestria'] = maestria
                        filtro = filtro & Q(inscripcioncohorte__cohortes__maestriaadmision__id=maestria)
                        cohortes = CohorteMaestria.objects.filter(maestriaadmision__id=maestria)
                        url_vars += "&maestria={}".format(maestria)

                    if cohorte > 0:
                        data['cohorte'] = cohorte
                        filtro = filtro & Q(inscripcioncohorte__cohortes__id=cohorte)
                        url_vars += "&cohorte={}".format(cohorte)

                    if asesorcomercial > 0:
                        data['asesor'] = asesorcomercial
                        filtro = filtro & Q(asesor=asesorcomercial)
                        url_vars += "&asesor={}".format(asesorcomercial)

                    if pais > 0:
                        data['pais'] = pais
                        filtro = filtro & Q(inscripcioncohorte__inscripcionaspirante__persona__pais__id=pais)
                        provincias = Provincia.objects.filter(pais__id=pais)
                        url_vars += "&pais={}".format(pais)

                    if provincia > 0:
                        data['provincia'] = provincia
                        filtro = filtro & Q(inscripcioncohorte__inscripcionaspirante__persona__provincia__id=provincia)
                        cantones = Canton.objects.filter(provincia__id=provincia)
                        url_vars += "&provincia={}".format(provincia)

                    if canton > 0:
                        data['canton'] = canton
                        filtro = filtro & Q(inscripcioncohorte__inscripcionaspirante__persona__canton__id=canton)
                        url_vars += "&canton={}".format(canton)

                    if desde and hasta:
                        data['desde'] = desde
                        data['hasta'] = hasta
                        url_vars += "&desde={}".format(desde)
                        url_vars += "&hasta={}".format(hasta)
                        filtro = filtro & Q(fecha__range=(desde, hasta))

                    elif desde:
                        data['desde'] = desde
                        url_vars += "&desde={}".format(desde)
                        filtro = filtro & Q(fecha__gte=desde)
                    elif hasta:
                        data['hasta'] = hasta
                        url_vars += "&hasta={}".format(hasta)
                        filtro = filtro & Q(fecha__lte=hasta)

                    if rol > 0:
                        data['idrol'] = rol
                        if rol == 2 or rol == 3 or rol == 4:
                            filtro = filtro & Q(inscripcioncohorte__asesor__rolgrupo=rol)
                        else:
                            filtro = filtro & Q(inscripcioncohorte__asesor__rol__id=rol)
                        url_vars += "&rol={}".format(rol)

                    if maestria == 0 and cohorte == 0 and asesorcomercial == 0 and pais == 0 and provincia == 0 and canton == 0 and desde == '' and hasta == '':
                        leads = []
                        asesores = []
                    else:
                        leads = VentasProgramaMaestria.objects.filter(filtro).exclude(valida=False).order_by('-id')
                        listado_reportadas.append(leads.values_list('inscripcioncohorte__id', flat=True))
                        tot = leads.count()
                        asesores = leads.values_list('asesor__id', flat=True).order_by('asesor__id').distinct().exclude(asesor__isnull=True)

                    dicc = {}
                    for ase in asesores:
                        asesor = AsesorComercial.objects.get(pk=ase)
                        qty2 = VentasProgramaMaestria.objects.filter(inscripcioncohorte__id__in=listado_reportadas, asesor=asesor).exclude(valida=False).count()
                        # ventas_ase.append(qty2)
                        nombreasesor = asesor.persona.apellido1 + ' ' + asesor.persona.apellido2 + ' ' + asesor.persona.nombres
                        # listado_asesor.append(nombreasesor)
                        dicc = {'asesor': nombreasesor, 'cantidad': qty2, 'id': asesor.id}
                        listado_asesor.append(dicc)
                    newlist = sorted(listado_asesor, key=lambda d: d['cantidad'], reverse=True)
                    lis_ase = [d['asesor'] for d in newlist]
                    lis_can = [d['cantidad'] for d in newlist]
                    lis_id = [d['id'] for d in newlist]

                    if len(lis_ase) > 0 and len(lis_can):
                        primero = lis_ase[0]
                        segundo = lis_ase[1] if len(lis_ase) > 1 else ''
                        tercero = lis_ase[2] if len(lis_ase) > 2 else ''

                        prican = lis_can[0]
                        segcan = lis_can[1] if len(lis_can) > 1 else ''
                        tercan = lis_can[2] if len(lis_can) > 2 else ''

                        first = AsesorComercial.objects.get(pk=lis_id[0])
                        second = AsesorComercial.objects.get(pk=lis_id[1]) if len(lis_id) > 1 else ''
                        third = AsesorComercial.objects.get(pk=lis_id[2]) if len(lis_id) > 2 else ''

                        # lista_provincias.append('SIN PROVINCIA')
                    data['listado_reportadas'] = listado_reportadas
                    data['labelasesor'] = lis_ase
                    data['asesores'] = AsesorComercial.objects.filter(id__in=asesores)
                    data['cant_ventas_ase'] = lis_can
                    # data['total_ventas'] = VentasProgramaMaestria.objects.filter(inscripcioncohorte__id__in=listado_reportadas).exclude(valida=False).count()
                    data['maestrialist'] = MaestriasAdmision.objects.filter(status=True, carrera__coordinacion__id=7)
                    data['cohorteslist'] = cohortes
                    data['asesorlist'] = AsesorComercial.objects.filter(status=True)
                    data['paislist'] = Pais.objects.filter(status=True)
                    data['provincialist'] = provincias
                    data['cantonlist'] = cantones
                    data['url_vars'] = url_vars

                    data['primero'] = primero
                    data['prican'] = prican
                    data['segundo'] = segundo
                    data['segcan'] = segcan
                    data['tercero'] = tercero
                    data['tercan'] = tercan
                    data['first'] = first
                    data['second'] = second
                    data['third'] = third
                    data['roles'] = RolAsesor.objects.filter(status=True, id__in=[1, 6])
                    data['total_ventas'] = tot
                    return render(request, "estadisticas/mapeoventasasesor.html", data)
                except Exception as ex:
                    pass

            elif action == 'mapeoventasrecuadado':
                try:
                    from sagest.models import Pago
                    data['title'] = u'Mapeo de dinero recaudado por asesores comerciales'
                    listado_reportadas = []
                    # ventas_ase = []
                    listado_asesor = []
                    tot = 0

                    maestria = cohorte = asesorcomercial = pais = provincia = canton = anio = rol = 0
                    desde = hasta = ''

                    filtro = Q(status=True)
                    if 'maestria' in request.GET:
                        maestria = int(request.GET['maestria'])
                    if 'cohorte' in request.GET:
                        cohorte = int(request.GET['cohorte'])
                    if 'asesor' in request.GET:
                        asesorcomercial = int(request.GET['asesor'])
                    if 'pais' in request.GET:
                        pais = int(request.GET['pais'])
                    if 'provincia' in request.GET:
                        provincia = int(request.GET['provincia'])
                    if 'canton' in request.GET:
                        canton = int(request.GET['canton'])
                    if 'desde' in request.GET:
                        desde = request.GET['desde']
                    if 'hasta' in request.GET:
                        hasta = request.GET['hasta']
                    if 'rol' in request.GET:
                        rol = int(request.GET['rol'])

                    url_vars = ' '
                    cohortes = provincias = cantones= None
                    primero = segundo = tercero = ''
                    prican = segcan = tercan = 0
                    first = second = third = None

                    if maestria > 0:
                        data['maestria'] = maestria
                        filtro = filtro & Q(inscripcioncohorte__cohortes__maestriaadmision__id=maestria)
                        cohortes = CohorteMaestria.objects.filter(maestriaadmision__id=maestria)
                        url_vars += "&maestria={}".format(maestria)

                    if cohorte > 0:
                        data['cohorte'] = cohorte
                        filtro = filtro & Q(inscripcioncohorte__cohortes__id=cohorte)
                        url_vars += "&cohorte={}".format(cohorte)

                    if asesorcomercial > 0:
                        data['asesor'] = asesorcomercial
                        filtro = filtro & Q(asesor=asesorcomercial)
                        url_vars += "&asesor={}".format(asesorcomercial)

                    if pais > 0:
                        data['pais'] = pais
                        filtro = filtro & Q(inscripcioncohorte__inscripcionaspirante__persona__pais__id=pais)
                        provincias = Provincia.objects.filter(pais__id=pais)
                        url_vars += "&pais={}".format(pais)

                    if provincia > 0:
                        data['provincia'] = provincia
                        filtro = filtro & Q(inscripcioncohorte__inscripcionaspirante__persona__provincia__id=provincia)
                        cantones = Canton.objects.filter(provincia__id=provincia)
                        url_vars += "&provincia={}".format(provincia)

                    if canton > 0:
                        data['canton'] = canton
                        filtro = filtro & Q(inscripcioncohorte__inscripcionaspirante__persona__canton__id=canton)
                        url_vars += "&canton={}".format(canton)

                    if desde and hasta:
                        data['desde'] = desde
                        data['hasta'] = hasta
                        url_vars += "&desde={}".format(desde)
                        url_vars += "&hasta={}".format(hasta)
                        filtro = filtro & Q(fecha__range=(desde, hasta))

                    elif desde:
                        data['desde'] = desde
                        url_vars += "&desde={}".format(desde)
                        filtro = filtro & Q(fecha__gte=desde)
                    elif hasta:
                        data['hasta'] = hasta
                        url_vars += "&hasta={}".format(hasta)
                        filtro = filtro & Q(fecha__lte=hasta)

                    if rol > 0:
                        data['idrol'] = rol
                        if rol == 2 or rol == 3 or rol == 4:
                            filtro = filtro & Q(inscripcioncohorte__asesor__rolgrupo=rol)
                        else:
                            filtro = filtro & Q(inscripcioncohorte__asesor__rol__id=rol)
                        url_vars += "&rol={}".format(rol)

                    if maestria == 0 and cohorte == 0 and asesorcomercial == 0 and pais == 0 and provincia == 0 and canton == 0 and desde == '' and hasta == '':
                        leads = []
                        asesores = []
                    else:
                        leads = VentasProgramaMaestria.objects.filter(filtro).exclude(valida=False).order_by('-id')
                        listado_reportadas.append(leads.values_list('inscripcioncohorte__id', flat=True))
                        # tot = leads.count()
                        idst = Rubro.objects.filter(status=True, inscripcion__id__in=listado_reportadas).values_list('id', flat=True)
                        tot = null_to_decimal(Pago.objects.filter(status=True, rubro__id__in=idst).aggregate(total=Sum('valortotal'))['total'], 2)
                        asesores = leads.values_list('asesor__id', flat=True).order_by('asesor__id').distinct().exclude(asesor__isnull=True)

                    dicc = {}
                    for ase in asesores:
                        asesor = AsesorComercial.objects.get(pk=ase)
                        ids = Rubro.objects.filter(status=True, inscripcion__id__in=listado_reportadas, inscripcion__asesor=asesor).values_list('id', flat=True)
                        qty2 = null_to_decimal(Pago.objects.filter(status=True, rubro__id__in=ids).aggregate(total=Sum('valortotal'))['total'], 2)
                        # qty2 = Decimal(null_to_decimal(Pago.objects.filter(status=True, rubro__id__in=ids).aggregate(total=Sum('valortotal'))['total'], 2)).quantize(Decimal('.01'))
                        # qty2 = VentasProgramaMaestria.objects.filter(inscripcioncohorte__id__in=listado_reportadas, asesor=asesor).exclude(valida=False).count()
                        # ventas_ase.append(qty2)
                        nombreasesor = asesor.persona.apellido1 + ' ' + asesor.persona.apellido2 + ' ' + asesor.persona.nombres + ' - ' + str(asesor.ventas_obtenidas_fecha(listado_reportadas))
                        # listado_asesor.append(nombreasesor)
                        dicc = {'asesor': nombreasesor, 'cantidad': qty2, 'id': asesor.id}
                        listado_asesor.append(dicc)
                    newlist = sorted(listado_asesor, key=lambda d: d['cantidad'], reverse=True)
                    lis_ase = [d['asesor'] for d in newlist]
                    lis_can = [d['cantidad'] for d in newlist]
                    lis_id = [d['id'] for d in newlist]

                    if len(lis_ase) > 0 and len(lis_can):
                        primero = lis_ase[0]
                        segundo = lis_ase[1] if len(lis_ase) > 1 else ''
                        tercero = lis_ase[2] if len(lis_ase) > 2 else ''

                        prican = Decimal(lis_can[0]).quantize(Decimal('.01'))
                        segcan = Decimal(lis_can[1]).quantize(Decimal('.01')) if len(lis_can) > 1 else ''
                        tercan = Decimal(lis_can[2]).quantize(Decimal('.01')) if len(lis_can) > 2 else ''

                        first = AsesorComercial.objects.get(pk=lis_id[0])
                        second = AsesorComercial.objects.get(pk=lis_id[1]) if len(lis_id) > 1 else ''
                        third = AsesorComercial.objects.get(pk=lis_id[2]) if len(lis_id) > 2 else ''

                        # lista_provincias.append('SIN PROVINCIA')
                    data['listado_reportadas'] = listado_reportadas
                    data['labelasesor'] = lis_ase
                    data['asesores'] = AsesorComercial.objects.filter(id__in=asesores)
                    data['cant_ventas_ase'] = lis_can
                    # data['total_ventas'] = VentasProgramaMaestria.objects.filter(inscripcioncohorte__id__in=listado_reportadas).exclude(valida=False).count()
                    data['maestrialist'] = MaestriasAdmision.objects.filter(status=True, carrera__coordinacion__id=7)
                    data['cohorteslist'] = cohortes
                    data['asesorlist'] = AsesorComercial.objects.filter(status=True)
                    data['paislist'] = Pais.objects.filter(status=True)
                    data['provincialist'] = provincias
                    data['cantonlist'] = cantones
                    data['url_vars'] = url_vars

                    data['primero'] = primero
                    data['prican'] = prican
                    data['segundo'] = segundo
                    data['segcan'] = segcan
                    data['tercero'] = tercero
                    data['tercan'] = tercan
                    data['first'] = first
                    data['second'] = second
                    data['third'] = third
                    data['roles'] = RolAsesor.objects.filter(status=True, id__in=[1, 6])
                    data['total_ventas'] = tot
                    return render(request, "estadisticas/mapeoventasrecuadado.html", data)
                except Exception as ex:
                    pass

            elif action == 'mapeoventascohortes':
                try:
                    data['title'] = u'Mapeo de ventas por Cohortes de maestría'
                    listado_reportadas = []
                    # ventas_ase = []
                    listado_cohortes = []
                    tot = 0

                    maestria = cohorte = asesorcomercial = pais = provincia = canton = 0
                    desde = hasta = ''

                    filtro = Q(status=True)
                    if 'maestria' in request.GET:
                        maestria = int(request.GET['maestria'])
                    if 'cohorte' in request.GET:
                        cohorte = int(request.GET['cohorte'])
                    if 'asesor' in request.GET:
                        asesorcomercial = int(request.GET['asesor'])
                    if 'pais' in request.GET:
                        pais = int(request.GET['pais'])
                    if 'provincia' in request.GET:
                        provincia = int(request.GET['provincia'])
                    if 'canton' in request.GET:
                        canton = int(request.GET['canton'])
                    if 'desde' in request.GET:
                        desde = request.GET['desde']
                    if 'hasta' in request.GET:
                        hasta = request.GET['hasta']

                    url_vars = ' '
                    cohortes = provincias = cantones= None
                    primero = segundo = tercero = ''
                    prican = segcan = tercan = 0
                    first = second = third = None

                    if maestria > 0:
                        data['maestria'] = maestria
                        filtro = filtro & Q(inscripcioncohorte__cohortes__maestriaadmision__id=maestria)
                        cohortes = CohorteMaestria.objects.filter(maestriaadmision__id=maestria)
                        url_vars += "&maestria={}".format(maestria)

                    if cohorte > 0:
                        data['cohorte'] = cohorte
                        filtro = filtro & Q(inscripcioncohorte__cohortes__id=cohorte)
                        url_vars += "&cohorte={}".format(cohorte)

                    if asesorcomercial > 0:
                        data['asesor'] = asesorcomercial
                        filtro = filtro & Q(asesor=asesorcomercial)
                        url_vars += "&asesor={}".format(asesorcomercial)

                    if pais > 0:
                        data['pais'] = pais
                        filtro = filtro & Q(inscripcioncohorte__inscripcionaspirante__persona__pais__id=pais)
                        provincias = Provincia.objects.filter(pais__id=pais)
                        url_vars += "&pais={}".format(pais)

                    if provincia > 0:
                        data['provincia'] = provincia
                        filtro = filtro & Q(inscripcioncohorte__inscripcionaspirante__persona__provincia__id=provincia)
                        cantones = Canton.objects.filter(provincia__id=provincia)
                        url_vars += "&provincia={}".format(provincia)

                    if canton > 0:
                        data['canton'] = canton
                        filtro = filtro & Q(inscripcioncohorte__inscripcionaspirante__persona__canton__id=canton)
                        url_vars += "&canton={}".format(canton)

                    if desde and hasta:
                        data['desde'] = desde
                        data['hasta'] = hasta
                        url_vars += "&desde={}".format(desde)
                        url_vars += "&hasta={}".format(hasta)
                        filtro = filtro & Q(fecha__range=(desde, hasta))

                    elif desde:
                        data['desde'] = desde
                        url_vars += "&desde={}".format(desde)
                        filtro = filtro & Q(fecha__gte=desde)
                    elif hasta:
                        data['hasta'] = hasta
                        url_vars += "&hasta={}".format(hasta)
                        filtro = filtro & Q(fecha__lte=hasta)

                    if maestria == 0 and cohorte == 0 and asesorcomercial == 0 and pais == 0 and provincia == 0 and canton == 0 and desde == '' and hasta == '':
                        leads = []
                        cohortes_ventas = []
                    else:
                        leads = VentasProgramaMaestria.objects.filter(filtro).exclude(valida=False).order_by('-id')
                        listado_reportadas.append(leads.values_list('inscripcioncohorte__id', flat=True))
                        tot = leads.count()
                        cohortes_ventas = leads.values_list('inscripcioncohorte__cohortes__id', flat=True).order_by('inscripcioncohorte__cohortes__id').distinct()

                    dicc = {}
                    for coho in cohortes_ventas:
                        cohortem = CohorteMaestria.objects.get(pk=coho)
                        qty2 = VentasProgramaMaestria.objects.filter(inscripcioncohorte__id__in=listado_reportadas, inscripcioncohorte__cohortes=cohortem).exclude(valida=False).count()
                        nombrecohorte = cohortem.maestriaadmision.descripcion + ' - ' + cohortem.descripcion
                        dicc = {'cohorte': nombrecohorte, 'cantidad': qty2}
                        listado_cohortes.append(dicc)
                    newlist = sorted(listado_cohortes, key=lambda d: d['cantidad'], reverse=True)
                    lis_coho = [d['cohorte'] for d in newlist]
                    lis_can = [d['cantidad'] for d in newlist]

                    data['listado_reportadas'] = listado_reportadas
                    data['labelcohorte'] = lis_coho
                    data['cohortesmaestria'] = CohorteMaestria.objects.filter(id__in=cohortes_ventas)
                    data['cant_ventas_coho'] = lis_can
                    # data['total_ventas'] = InscripcionCohorte.objects.filter(id__in=listado_reportadas).count()
                    data['maestrialist'] = MaestriasAdmision.objects.filter(status=True, carrera__coordinacion__id=7)
                    data['cohorteslist'] = cohortes
                    data['asesorlist'] = AsesorComercial.objects.filter(status=True)
                    data['paislist'] = Pais.objects.filter(status=True)
                    data['provincialist'] = provincias
                    data['cantonlist'] = cantones
                    data['url_vars'] = url_vars

                    data['total_ventas'] = tot
                    return render(request, "estadisticas/mapeoventascohortes.html", data)
                except Exception as ex:
                    pass

            elif action == 'estadisticasgeneralevalpos':
                try:
                    data['title'] = u'Estadísticas generales de evalución Posgrado'

                    filtro = Q(status=True, materia__nivel__periodo__tipo__id=3, tipoprofesor__id=11,
                               materia__fin__lt=datetime.now().date())

                    desde = hasta = ''
                    url_vars = ' '

                    if 'desde' in request.GET:
                        desde = request.GET['desde']
                    if 'hasta' in request.GET:
                        hasta = request.GET['hasta']

                    if desde and hasta:
                        data['desde'] = desde
                        data['hasta'] = hasta
                        filtro = filtro & Q(materia__fin__range=(desde, hasta))
                        url_vars += "&desde={}".format(desde)
                        url_vars += "&hasta={}".format(hasta)

                    elif desde:
                        data['desde'] = desde
                        filtro = filtro & Q(materia__fin__gte=desde)
                        url_vars += "&desde={}".format(hasta)

                    elif hasta:
                        data['hasta'] = hasta
                        filtro = filtro & Q(materia__fin__lte=hasta)
                        url_vars += "&hasta={}".format(hasta)


                    if desde == '' and hasta == '':
                        cvnt_beta = cvnt_omega = cvnt_alfa = total = suma =0
                        cvnt_beta_edu = cvnt_omega_edu = cvnt_alfa_edu = total_edu = suma_edu = 0
                        cvnt_beta_neg = cvnt_omega_neg = cvnt_alfa_neg = total_neg = suma_neg = 0
                        cvnt_beta_sal = cvnt_omega_sal = cvnt_alfa_sal = total_sal = suma_sal = 0
                    else:
                        eModulos = ProfesorMateria.objects.filter(filtro).order_by('-id')

                        cvnt_beta = cvnt_omega = cvnt_alfa = total = suma = 0
                        cvnt_beta_edu = cvnt_omega_edu = cvnt_alfa_edu = total_edu = suma_edu = 0
                        cvnt_beta_neg = cvnt_omega_neg = cvnt_alfa_neg = total_neg = suma_neg = 0
                        cvnt_beta_sal = cvnt_omega_sal = cvnt_alfa_sal = total_sal = suma_sal = 0
                        for eModulo in eModulos:
                            eProfesor = eModulo.profesor
                            distributivo = eProfesor.distributivohoraseval(eModulo.materia.nivel.periodo)
                            if distributivo and distributivo.resumen_evaluacion_acreditacion():
                                resumen = distributivo.resumen_evaluacion_acreditacion()
                                if resumen.resultado_docencia and resumen.resultado_docencia > 0:
                                    if cincoacien(resumen.resultado_docencia) > 0 and cincoacien(resumen.resultado_docencia) <= 70:
                                        cvnt_beta += 1
                                        if eModulo.materia.asignaturamalla.malla.carrera.escuelaposgrado.id == 1:
                                            cvnt_beta_sal += 1
                                            suma_sal += cincoacien(resumen.resultado_docencia)
                                        elif eModulo.materia.asignaturamalla.malla.carrera.escuelaposgrado.id == 2:
                                            cvnt_beta_edu += 1
                                            suma_edu += cincoacien(resumen.resultado_docencia)
                                        elif eModulo.materia.asignaturamalla.malla.carrera.escuelaposgrado.id == 3:
                                            cvnt_beta_neg += 1
                                            suma_neg += cincoacien(resumen.resultado_docencia)
                                    elif cincoacien(resumen.resultado_docencia) > 70 and cincoacien(resumen.resultado_docencia) <= 90:
                                        cvnt_omega += 1
                                        if eModulo.materia.asignaturamalla.malla.carrera.escuelaposgrado.id == 1:
                                            cvnt_omega_sal += 1
                                            suma_sal += cincoacien(resumen.resultado_docencia)
                                        elif eModulo.materia.asignaturamalla.malla.carrera.escuelaposgrado.id == 2:
                                            cvnt_omega_edu += 1
                                            suma_edu += cincoacien(resumen.resultado_docencia)
                                        elif eModulo.materia.asignaturamalla.malla.carrera.escuelaposgrado.id == 3:
                                            cvnt_omega_neg += 1
                                            suma_neg += cincoacien(resumen.resultado_docencia)
                                    elif cincoacien(resumen.resultado_docencia) > 90:
                                        cvnt_alfa += 1
                                        if eModulo.materia.asignaturamalla.malla.carrera.escuelaposgrado.id == 1:
                                            cvnt_alfa_sal += 1
                                            suma_sal += cincoacien(resumen.resultado_docencia)
                                        elif eModulo.materia.asignaturamalla.malla.carrera.escuelaposgrado.id == 2:
                                            cvnt_alfa_edu += 1
                                            suma_edu += cincoacien(resumen.resultado_docencia)
                                        elif eModulo.materia.asignaturamalla.malla.carrera.escuelaposgrado.id == 3:
                                            cvnt_alfa_neg += 1
                                            suma_neg += cincoacien(resumen.resultado_docencia)
                                    total += 1
                                    suma += cincoacien(resumen.resultado_docencia)
                    dicc1 = dicc2 = dicc3 = {}
                    dicc1 = {'etiqueta': '0-70', 'cantidad': cvnt_beta, 'orden': 1}
                    dicc2 = {'etiqueta': '70-90', 'cantidad': cvnt_omega, 'orden': 2}
                    dicc3 = {'etiqueta': '90-100', 'cantidad': cvnt_alfa, 'orden': 3}
                    listado = []
                    listado.append(dicc1)
                    listado.append(dicc2)
                    listado.append(dicc3)

                    newlist = sorted(listado, key=lambda d: d['orden'], reverse=True)
                    eti = [d['etiqueta'] for d in newlist]
                    canti = [d['cantidad'] for d in newlist]

                    data['canti'] = canti
                    data['etiquetas'] = eti
                    data['total_eval'] = total
                    if suma > 0 and total > 0:
                        promedio = suma/total
                    else:
                        promedio = 0
                    data['promedio_general'] = round(promedio, 2)

                    dicc4 = {'etiqueta': '0-70', 'cantidad': cvnt_beta_sal, 'orden': 1}
                    dicc5 = {'etiqueta': '70-90', 'cantidad': cvnt_omega_sal, 'orden': 2}
                    dicc6 = {'etiqueta': '90-100', 'cantidad': cvnt_alfa_sal, 'orden': 3}
                    listado_salud = []
                    listado_salud.append(dicc4)
                    listado_salud.append(dicc5)
                    listado_salud.append(dicc6)

                    newlist_sal = sorted(listado_salud, key=lambda d: d['orden'], reverse=True)
                    eti_sal = [d['etiqueta'] for d in newlist_sal]
                    canti_sal = [d['cantidad'] for d in newlist_sal]
                    total_sal = cvnt_beta_sal + cvnt_omega_sal + cvnt_alfa_sal
                    data['canti_sal'] = canti_sal
                    data['eti_sal'] = eti_sal
                    data['total_sal'] = total_sal
                    if suma_sal > 0 and total_sal > 0:
                        promedio_sal = suma_sal/total_sal
                    else:
                        promedio_sal = 0
                    data['promedio_sal'] = round(promedio_sal, 2)

                    dicc7 = {'etiqueta': '0-70', 'cantidad': cvnt_beta_edu, 'orden': 1}
                    dicc8 = {'etiqueta': '70-90', 'cantidad': cvnt_omega_edu, 'orden': 2}
                    dicc9 = {'etiqueta': '90-100', 'cantidad': cvnt_alfa_edu, 'orden': 3}
                    listado_edu = []
                    listado_edu.append(dicc7)
                    listado_edu.append(dicc8)
                    listado_edu.append(dicc9)

                    newlist_edu = sorted(listado_edu, key=lambda d: d['orden'], reverse=True)
                    eti_edu = [d['etiqueta'] for d in newlist_edu]
                    canti_edu = [d['cantidad'] for d in newlist_edu]
                    total_edu = cvnt_beta_edu + cvnt_omega_edu + cvnt_alfa_edu
                    data['canti_edu'] = canti_edu
                    data['eti_edu'] = eti_edu
                    data['total_edu'] = total_edu
                    if suma_edu > 0 and total_edu > 0:
                        promedio_edu = suma_edu/total_edu
                    else:
                        promedio_edu = 0
                    data['promedio_edu'] = round(promedio_edu, 2)

                    dicc10 = {'etiqueta': '0-70', 'cantidad': cvnt_beta_neg, 'orden': 1}
                    dicc11 = {'etiqueta': '70-90', 'cantidad': cvnt_omega_neg, 'orden': 2}
                    dicc12 = {'etiqueta': '90-100', 'cantidad': cvnt_alfa_neg, 'orden': 3}
                    listado_neg = []
                    listado_neg.append(dicc10)
                    listado_neg.append(dicc11)
                    listado_neg.append(dicc12)

                    newlist_neg = sorted(listado_neg, key=lambda d: d['orden'], reverse=True)
                    eti_neg = [d['etiqueta'] for d in newlist_neg]
                    canti_neg = [d['cantidad'] for d in newlist_neg]
                    total_neg = cvnt_beta_neg + cvnt_omega_neg + cvnt_alfa_neg
                    data['canti_neg'] = canti_neg
                    data['eti_neg'] = eti_neg
                    data['total_neg'] = total_neg
                    if suma_neg > 0 and total_neg > 0:
                        promedio_neg = suma_neg/total_neg
                    else:
                        promedio_neg = 0
                    data['promedio_neg'] = round(promedio_neg, 2)
                    return render(request, "estadisticas/generalevalpos.html", data)
                except Exception as ex:
                    pass

            elif action == 'estadisticasgeneralevalpos_sal':
                try:
                    data['title'] = u'Estadísticas generales de Escuela de Salud'

                    filtro = Q(status=True, materia__nivel__periodo__tipo__id=3, tipoprofesor__id=11, materia__asignaturamalla__malla__carrera__escuelaposgrado__id=1,
                               materia__fin__lt=datetime.now().date())

                    desde = hasta = ''
                    url_vars = ' '

                    if 'desde' in request.GET:
                        desde = request.GET['desde']
                    if 'hasta' in request.GET:
                        hasta = request.GET['hasta']

                    if desde and hasta:
                        data['desde'] = desde
                        data['hasta'] = hasta
                        filtro = filtro & Q(materia__fin__range=(desde, hasta))
                        url_vars += "&desde={}".format(desde)
                        url_vars += "&hasta={}".format(hasta)

                    elif desde:
                        data['desde'] = desde
                        filtro = filtro & Q(materia__fin__gte=desde)
                        url_vars += "&desde={}".format(hasta)

                    elif hasta:
                        data['hasta'] = hasta
                        filtro = filtro & Q(materia__fin__lte=hasta)
                        url_vars += "&hasta={}".format(hasta)

                    lista_c = []
                    if desde == '' and hasta == '':
                        cvnt_beta_sal = cvnt_omega_sal = cvnt_alfa_sal = total_sal = suma_sal = 0
                    else:
                        eModulos = ProfesorMateria.objects.filter(filtro).order_by('-id')
                        cvnt_beta_sal = cvnt_omega_sal = cvnt_alfa_sal = total_sal = suma_sal = 0
                        for eModulo in eModulos:
                            eProfesor = eModulo.profesor
                            distributivo = eProfesor.distributivohoraseval(eModulo.materia.nivel.periodo)
                            if distributivo and distributivo.resumen_evaluacion_acreditacion():
                                resumen = distributivo.resumen_evaluacion_acreditacion()
                                if resumen.resultado_docencia and resumen.resultado_docencia > 0:
                                    if cincoacien(resumen.resultado_docencia) > 0 and cincoacien(resumen.resultado_docencia) <= 70:
                                        if eModulo.materia.asignaturamalla.malla.carrera.escuelaposgrado.id == 1:
                                            cvnt_beta_sal += 1
                                            suma_sal += cincoacien(resumen.resultado_docencia)
                                    elif cincoacien(resumen.resultado_docencia) > 70 and cincoacien(resumen.resultado_docencia) <= 90:
                                        if eModulo.materia.asignaturamalla.malla.carrera.escuelaposgrado.id == 1:
                                            cvnt_omega_sal += 1
                                            suma_sal += cincoacien(resumen.resultado_docencia)
                                    elif cincoacien(resumen.resultado_docencia) > 90:
                                        if eModulo.materia.asignaturamalla.malla.carrera.escuelaposgrado.id == 1:
                                            cvnt_alfa_sal += 1
                                            suma_sal += cincoacien(resumen.resultado_docencia)

                                    if eModulo.materia.asignaturamalla.malla.carrera.id not in lista_c:
                                        lista_c.append(eModulo.materia.asignaturamalla.malla.carrera.id)

                    dicc4 = {'etiqueta': '0-70', 'cantidad': cvnt_beta_sal, 'orden': 1}
                    dicc5 = {'etiqueta': '70-90', 'cantidad': cvnt_omega_sal, 'orden': 2}
                    dicc6 = {'etiqueta': '90-100', 'cantidad': cvnt_alfa_sal, 'orden': 3}
                    listado_salud = []
                    listado_salud.append(dicc4)
                    listado_salud.append(dicc5)
                    listado_salud.append(dicc6)

                    newlist_sal = sorted(listado_salud, key=lambda d: d['orden'], reverse=True)
                    eti_sal = [d['etiqueta'] for d in newlist_sal]
                    canti_sal = [d['cantidad'] for d in newlist_sal]
                    total_sal = cvnt_beta_sal + cvnt_omega_sal + cvnt_alfa_sal
                    data['canti_sal'] = canti_sal
                    data['eti_sal'] = eti_sal
                    data['total_sal'] = total_sal
                    if suma_sal > 0 and total_sal > 0:
                        promedio_sal = suma_sal/total_sal
                    else:
                        promedio_sal = 0
                    data['promedio_sal'] = round(promedio_sal, 2)
                    data['eCarreras'] = Carrera.objects.filter(status=True, escuelaposgrado__id=1, id__in=lista_c)
                    return render(request, "estadisticas/salu_evalpos.html", data)
                except Exception as ex:
                    pass

            elif action == 'estadisticasgeneralevalpos_neg':
                try:
                    data['title'] = u'Estadísticas generales de Escuela de Negocios'

                    filtro = Q(status=True, materia__nivel__periodo__tipo__id=3, tipoprofesor__id=11, materia__asignaturamalla__malla__carrera__escuelaposgrado__id=3,
                               materia__fin__lt=datetime.now().date())

                    desde = hasta = ''
                    url_vars = ' '

                    if 'desde' in request.GET:
                        desde = request.GET['desde']
                    if 'hasta' in request.GET:
                        hasta = request.GET['hasta']

                    if desde and hasta:
                        data['desde'] = desde
                        data['hasta'] = hasta
                        filtro = filtro & Q(materia__fin__range=(desde, hasta))
                        url_vars += "&desde={}".format(desde)
                        url_vars += "&hasta={}".format(hasta)

                    elif desde:
                        data['desde'] = desde
                        filtro = filtro & Q(materia__fin__gte=desde)
                        url_vars += "&desde={}".format(hasta)

                    elif hasta:
                        data['hasta'] = hasta
                        filtro = filtro & Q(materia__fin__lte=hasta)
                        url_vars += "&hasta={}".format(hasta)

                    lista_c = []
                    if desde == '' and hasta == '':
                        cvnt_beta_neg = cvnt_omega_neg = cvnt_alfa_neg = total_neg = suma_neg = 0
                    else:
                        eModulos = ProfesorMateria.objects.filter(filtro).order_by('-id')
                        cvnt_beta_neg = cvnt_omega_neg = cvnt_alfa_neg = total_neg = suma_neg = 0
                        for eModulo in eModulos:
                            eProfesor = eModulo.profesor
                            distributivo = eProfesor.distributivohoraseval(eModulo.materia.nivel.periodo)
                            if distributivo and distributivo.resumen_evaluacion_acreditacion():
                                resumen = distributivo.resumen_evaluacion_acreditacion()
                                if resumen.resultado_docencia and resumen.resultado_docencia > 0:
                                    if cincoacien(resumen.resultado_docencia) > 0 and cincoacien(resumen.resultado_docencia) <= 70:
                                        if eModulo.materia.asignaturamalla.malla.carrera.escuelaposgrado.id == 3:
                                            cvnt_beta_neg += 1
                                            suma_neg += cincoacien(resumen.resultado_docencia)
                                    elif cincoacien(resumen.resultado_docencia) > 70 and cincoacien(resumen.resultado_docencia) <= 90:
                                        if eModulo.materia.asignaturamalla.malla.carrera.escuelaposgrado.id == 3:
                                            cvnt_omega_neg += 1
                                            suma_neg += cincoacien(resumen.resultado_docencia)
                                    elif cincoacien(resumen.resultado_docencia) > 90:
                                        if eModulo.materia.asignaturamalla.malla.carrera.escuelaposgrado.id == 3:
                                            cvnt_alfa_neg += 1
                                            suma_neg += cincoacien(resumen.resultado_docencia)

                                    if eModulo.materia.asignaturamalla.malla.carrera.id not in lista_c:
                                        lista_c.append(eModulo.materia.asignaturamalla.malla.carrera.id)

                    dicc4 = {'etiqueta': '0-70', 'cantidad': cvnt_beta_neg, 'orden': 1}
                    dicc5 = {'etiqueta': '70-90', 'cantidad': cvnt_omega_neg, 'orden': 2}
                    dicc6 = {'etiqueta': '90-100', 'cantidad': cvnt_alfa_neg, 'orden': 3}
                    listado_neg = []
                    listado_neg.append(dicc4)
                    listado_neg.append(dicc5)
                    listado_neg.append(dicc6)

                    newlist_neg = sorted(listado_neg, key=lambda d: d['orden'], reverse=True)
                    eti_neg = [d['etiqueta'] for d in newlist_neg]
                    canti_neg = [d['cantidad'] for d in newlist_neg]
                    total_neg = cvnt_beta_neg + cvnt_omega_neg + cvnt_alfa_neg
                    data['canti_neg'] = canti_neg
                    data['eti_neg'] = eti_neg
                    data['total_neg'] = total_neg
                    if suma_neg > 0 and total_neg > 0:
                        promedio_neg = suma_neg/total_neg
                    else:
                        promedio_neg = 0
                    data['promedio_neg'] = round(promedio_neg, 2)
                    data['eCarreras'] = Carrera.objects.filter(status=True, escuelaposgrado__id=3, id__in=lista_c)
                    return render(request, "estadisticas/neg_evalpos.html", data)
                except Exception as ex:
                    pass

            elif action == 'estadisticasgeneralevalpos_edu':
                try:
                    data['title'] = u'Estadísticas generales de Escuela de Educación'

                    filtro = Q(status=True, materia__nivel__periodo__tipo__id=3, tipoprofesor__id=11, materia__asignaturamalla__malla__carrera__escuelaposgrado__id=2,
                               materia__fin__lt=datetime.now().date())

                    desde = hasta = ''
                    url_vars = ' '

                    if 'desde' in request.GET:
                        desde = request.GET['desde']
                    if 'hasta' in request.GET:
                        hasta = request.GET['hasta']

                    if desde and hasta:
                        data['desde'] = desde
                        data['hasta'] = hasta
                        filtro = filtro & Q(materia__fin__range=(desde, hasta))
                        url_vars += "&desde={}".format(desde)
                        url_vars += "&hasta={}".format(hasta)

                    elif desde:
                        data['desde'] = desde
                        filtro = filtro & Q(materia__fin__gte=desde)
                        url_vars += "&desde={}".format(hasta)

                    elif hasta:
                        data['hasta'] = hasta
                        filtro = filtro & Q(materia__fin__lte=hasta)
                        url_vars += "&hasta={}".format(hasta)

                    lista_c = []
                    if desde == '' and hasta == '':
                        cvnt_beta_edu = cvnt_omega_edu = cvnt_alfa_edu = total_edu = suma_edu = 0
                    else:
                        eModulos = ProfesorMateria.objects.filter(filtro).order_by('-id')
                        cvnt_beta_edu = cvnt_omega_edu = cvnt_alfa_edu = total_edu = suma_edu = 0
                        for eModulo in eModulos:
                            eProfesor = eModulo.profesor
                            distributivo = eProfesor.distributivohoraseval(eModulo.materia.nivel.periodo)
                            if distributivo and distributivo.resumen_evaluacion_acreditacion():
                                resumen = distributivo.resumen_evaluacion_acreditacion()
                                if resumen.resultado_docencia and resumen.resultado_docencia > 0:
                                    if cincoacien(resumen.resultado_docencia) > 0 and cincoacien(resumen.resultado_docencia) <= 70:
                                        if eModulo.materia.asignaturamalla.malla.carrera.escuelaposgrado.id == 2:
                                            cvnt_beta_edu += 1
                                            suma_edu += cincoacien(resumen.resultado_docencia)
                                    elif cincoacien(resumen.resultado_docencia) > 70 and cincoacien(resumen.resultado_docencia) <= 90:
                                        if eModulo.materia.asignaturamalla.malla.carrera.escuelaposgrado.id == 2:
                                            cvnt_omega_edu += 1
                                            suma_edu += cincoacien(resumen.resultado_docencia)
                                    elif cincoacien(resumen.resultado_docencia) > 90:
                                        if eModulo.materia.asignaturamalla.malla.carrera.escuelaposgrado.id == 2:
                                            cvnt_alfa_edu += 1
                                            suma_edu += cincoacien(resumen.resultado_docencia)

                                    if eModulo.materia.asignaturamalla.malla.carrera.id not in lista_c:
                                        lista_c.append(eModulo.materia.asignaturamalla.malla.carrera.id)
                    dicc4 = {'etiqueta': '0-70', 'cantidad': cvnt_beta_edu, 'orden': 1}
                    dicc5 = {'etiqueta': '70-90', 'cantidad': cvnt_omega_edu, 'orden': 2}
                    dicc6 = {'etiqueta': '90-100', 'cantidad': cvnt_alfa_edu, 'orden': 3}
                    listado_edu = []
                    listado_edu.append(dicc4)
                    listado_edu.append(dicc5)
                    listado_edu.append(dicc6)

                    newlist_edu = sorted(listado_edu, key=lambda d: d['orden'], reverse=True)
                    eti_edu = [d['etiqueta'] for d in newlist_edu]
                    canti_edu = [d['cantidad'] for d in newlist_edu]
                    total_edu = cvnt_beta_edu + cvnt_omega_edu + cvnt_alfa_edu
                    data['canti_edu'] = canti_edu
                    data['eti_edu'] = eti_edu
                    data['total_edu'] = total_edu
                    if suma_edu > 0 and total_edu > 0:
                        promedio_edu = suma_edu/total_edu
                    else:
                        promedio_edu = 0
                    data['promedio_edu'] = round(promedio_edu, 2)
                    data['eCarreras'] = Carrera.objects.filter(status=True, escuelaposgrado__id=2, id__in=lista_c)
                    return render(request, "estadisticas/edu_evalpos.html", data)
                except Exception as ex:
                    pass

            elif action == 'descargarreportegráfica':
                try:
                    __author__ = 'Unemi'

                    output = io.BytesIO()
                    workbook = xlsxwriter.Workbook(output)
                    ws = workbook.add_worksheet('general')
                    ws.set_column(0, 0, 10)
                    ws.set_column(1, 1, 15)
                    ws.set_column(2, 5, 30)
                    ws.set_column(6, 6, 20)
                    ws.set_column(7, 7, 40)
                    ws.set_column(8, 9, 20)
                    ws.set_column(10, 40, 30)

                    formatotitulo_filtros = workbook.add_format(
                        {'bold': 1, 'text_wrap': True, 'border': 1, 'align': 'center', 'font_size': 14})

                    formatoceldacab = workbook.add_format(
                        {'align': 'center', 'bold': 1, 'border': 1, 'text_wrap': True, 'fg_color': '#1C3247', 'font_color': 'white'})

                    formatoceldaleft = workbook.add_format(
                        {'text_wrap': True, 'align': 'center', 'valign': 'vcenter', 'border': 1})

                    formatoceldaleft_red = workbook.add_format(
                        {'text_wrap': True, 'align': 'center', 'valign': 'vcenter', 'border': 1, 'fg_color': '#FF0000', 'font_color': 'black'})

                    formatoceldaleft_yelow = workbook.add_format(
                        {'text_wrap': True, 'align': 'center', 'valign': 'vcenter', 'border': 1, 'fg_color': '#FFFF00', 'font_color': 'black'})

                    formatoceldaleft_green = workbook.add_format(
                        {'text_wrap': True, 'align': 'center', 'valign': 'vcenter', 'border': 1, 'fg_color': '#92D050', 'font_color': 'black'})

                    desde = request.GET.get('desde', '')
                    hasta = request.GET.get('hasta', '')

                    ws.merge_range('A1:AO1', 'SEGUIMIENTO DE ESTADÍSTICAS', formatotitulo_filtros)

                    ws.write(1, 0, 'N°', formatoceldacab)
                    ws.write(1, 1, 'ID', formatoceldacab)
                    ws.write(1, 2, 'Escuela', formatoceldacab)
                    ws.write(1, 3, 'Maestría', formatoceldacab)
                    ws.write(1, 4, 'Cohorte', formatoceldacab)
                    ws.write(1, 5, 'Módulo', formatoceldacab)
                    ws.write(1, 6, 'Paralelo', formatoceldacab)
                    ws.write(1, 7, 'Docente', formatoceldacab)
                    ws.write(1, 8, 'Inicio', formatoceldacab)
                    ws.write(1, 9, 'Fin', formatoceldacab)
                    ws.write(1, 10, 'Inicio eval_hetero', formatoceldacab)
                    ws.write(1, 11, 'Fin eval_hetero', formatoceldacab)
                    ws.write(1, 12, 'Estado eval_hetero', formatoceldacab)
                    ws.write(1, 13, 'Total evaluados', formatoceldacab)
                    ws.write(1, 14, 'Total matriculados', formatoceldacab)

                    ws.write(1, 15, 'Inicio eval_auto', formatoceldacab)
                    ws.write(1, 16, 'Fin eval_auto', formatoceldacab)
                    ws.write(1, 17, 'Estado eval_auto', formatoceldacab)
                    ws.write(1, 18, '¿Realizó Autoevaluación?', formatoceldacab)

                    ws.write(1, 19, 'Inicio eval_dir', formatoceldacab)
                    ws.write(1, 20, 'Fin eval_dir', formatoceldacab)
                    ws.write(1, 21, 'Estado eval_dir', formatoceldacab)
                    ws.write(1, 22, '¿Evaluado por Director?', formatoceldacab)
                    ws.write(1, 23, '¿Evaluado por Coordinador?', formatoceldacab)

                    ws.write(1, 24, '¿Procesado?', formatoceldacab)
                    ws.write(1, 25, 'Fecha de procesado', formatoceldacab)
                    ws.write(1, 26, 'Hora de procesado', formatoceldacab)

                    ws.write(1, 27, 'eval_hetero', formatoceldacab)
                    ws.write(1, 28, 'eval_hetero (%)', formatoceldacab)
                    ws.write(1, 29, 'eval_auto', formatoceldacab)
                    ws.write(1, 30, 'eval_auto (%)', formatoceldacab)
                    ws.write(1, 31, 'eval_directivo', formatoceldacab)
                    ws.write(1, 32, 'eval_directivo (%)', formatoceldacab)
                    ws.write(1, 33, 'evalp_hetero', formatoceldacab)
                    ws.write(1, 34, 'evalp_hetero (%)', formatoceldacab)
                    ws.write(1, 35, 'evalp_auto', formatoceldacab)
                    ws.write(1, 36, 'evalp_auto (%)', formatoceldacab)
                    ws.write(1, 37, 'evalp_directivo', formatoceldacab)
                    ws.write(1, 38, 'evalp_directivo (%)', formatoceldacab)
                    ws.write(1, 39, 'total_eval_abs', formatoceldacab)
                    ws.write(1, 40, 'total_eval_rel', formatoceldacab)

                    filtro = Q(status=True, tipoprofesor__id=11, materia__nivel__periodo__tipo__id=3, materia__fin__lt=datetime.now().date())

                    if desde and hasta:
                        filtro = filtro & Q(materia__fin__range=(desde, hasta))
                    elif desde:
                        filtro = filtro & Q(materia__fin__gte=desde)
                    elif hasta:
                        filtro = filtro & Q(materia__fin__lte=hasta)

                    eProfesorMateriasPrueba = ProfesorMateria.objects.filter(filtro).order_by('-id')
                    lis = []
                    cvnt = 0
                    for eModulo in eProfesorMateriasPrueba:
                        eProfesor = eModulo.profesor
                        distributivo = eProfesor.distributivohoraseval(eModulo.materia.nivel.periodo)
                        if distributivo and distributivo.resumen_evaluacion_acreditacion():
                            resumen = distributivo.resumen_evaluacion_acreditacion()
                            if resumen.resultado_docencia and resumen.resultado_docencia > 0:
                                lis.append(eModulo.id)
                                cvnt += 1
                                print(f'Contador: {cvnt}')

                    eProfesorMaterias = ProfesorMateria.objects.filter(id__in=lis).order_by('-id')
                    filas_recorridas = 3
                    cont = 1
                    for eProfesorMateria in eProfesorMaterias:
                        eMatriculados = eProfesorMateria.materia.materiaasignada_set.values("id").filter(status=True, matricula__estado_matricula__in=[2, 3], matricula__retiradomatricula=False).exclude(retiramateria=True).count()
                        eDir = object_dir(eProfesorMateria)
                        eRes = object_proces(eProfesorMateria)

                        distributivo = eProfesorMateria.profesor.distributivohoraseval(eProfesorMateria.materia.nivel.periodo)
                        if distributivo:
                            resumen = distributivo.resumen_evaluacion_acreditacion() if distributivo is not None else None
                        else:
                            resumen = None

                        formato = formatoceldaleft
                        if cincoacien(resumen.resultado_docencia) > 0 and cincoacien(resumen.resultado_docencia) <= 70:
                            formato = formatoceldaleft_red
                        elif cincoacien(resumen.resultado_docencia) > 70 and cincoacien(resumen.resultado_docencia) <= 90:
                            formato = formatoceldaleft_yelow
                        elif cincoacien(resumen.resultado_docencia) > 90:
                            formato = formatoceldaleft_green

                        ws.write('A%s' % filas_recorridas, str(cont), formatoceldaleft)
                        ws.write('B%s' % filas_recorridas, str(eProfesorMateria.id), formatoceldaleft)
                        ws.write('C%s' % filas_recorridas, str(eProfesorMateria.materia.asignaturamalla.malla.carrera.escuelaposgrado.nombre if eProfesorMateria.materia.asignaturamalla.malla.carrera.escuelaposgrado else 'NO REGISTRA'), formatoceldaleft)
                        ws.write('D%s' % filas_recorridas, str(eProfesorMateria.materia.asignaturamalla.malla.carrera), formatoceldaleft)
                        ws.write('E%s' % filas_recorridas, str(eProfesorMateria.materia.nivel.periodo.cohorte_maestria()), formatoceldaleft)
                        ws.write('F%s' % filas_recorridas, str(eProfesorMateria.materia.asignaturamalla.asignatura.nombre), formatoceldaleft)
                        ws.write('G%s' % filas_recorridas, str(eProfesorMateria.materia.paralelo), formatoceldaleft)
                        ws.write('H%s' % filas_recorridas, str(eProfesorMateria.profesor.persona), formatoceldaleft)
                        ws.write('I%s' % filas_recorridas, str(eProfesorMateria.materia.inicio), formatoceldaleft)
                        ws.write('J%s' % filas_recorridas, str(eProfesorMateria.materia.fin), formatoceldaleft)

                        ws.write('K%s' % filas_recorridas, str(eProfesorMateria.materia.inicioeval if eProfesorMateria.materia.inicioeval else 'NO CONFIGURADO'), formatoceldaleft)
                        ws.write('L%s' % filas_recorridas, str(eProfesorMateria.materia.fineval if eProfesorMateria.materia.fineval else 'NO CONFIGURADO'), formatoceldaleft)
                        ws.write('M%s' % filas_recorridas, str(fecha_vencida(eProfesorMateria.materia, 1)), formatoceldaleft)
                        ws.write('N%s' % filas_recorridas, cantidad_evaluacion_docente(eProfesorMateria), formatoceldaleft)
                        ws.write('O%s' % filas_recorridas, eMatriculados, formatoceldaleft)

                        ws.write('P%s' % filas_recorridas, str(eProfesorMateria.materia.inicioevalauto if eProfesorMateria.materia.inicioevalauto else 'NO CONFIGURADO'), formatoceldaleft)
                        ws.write('Q%s' % filas_recorridas, str(eProfesorMateria.materia.finevalauto if eProfesorMateria.materia.finevalauto else 'NO CONFIGURADO'), formatoceldaleft)
                        ws.write('R%s' % filas_recorridas, str(fecha_vencida(eProfesorMateria.materia, 2)), formatoceldaleft)
                        ws.write('S%s' % filas_recorridas, str('SI' if cantidad_evaluacion_auto(eProfesorMateria) > 0 else 'NO'), formatoceldaleft)

                        ws.write('T%s' % filas_recorridas, str(eDir.inicio if eDir else 'NO CONFIGURADO'), formatoceldaleft)
                        ws.write('U%s' % filas_recorridas, str(eDir.fin if eDir else 'NO CONFIGURADO'), formatoceldaleft)
                        ws.write('V%s' % filas_recorridas, str(fecha_vencida(eProfesorMateria.materia, 3)), formatoceldaleft)
                        ws.write('W%s' % filas_recorridas, str('SI' if evaluo_director(eProfesorMateria) else 'NO'), formatoceldaleft)
                        ws.write('X%s' % filas_recorridas, str('SI' if evaluo_coordinador(eProfesorMateria) else 'NO'), formatoceldaleft)

                        ws.write('Y%s' % filas_recorridas, str('SI' if eRes else 'NO PROCESADO'), formatoceldaleft)
                        ws.write('Z%s' % filas_recorridas, str(eRes.fecha_creacion.date().strftime('%d-%m-%Y') if eRes else 'NO PROCESADO'), formatoceldaleft)
                        ws.write('AA%s' % filas_recorridas, str(eRes.fecha_creacion.time().strftime('%I:%M %p') if eRes else 'NO PROCESADO'), formatoceldaleft)

                        ws.write('AB%s' % filas_recorridas, resumen.promedio_docencia_hetero if resumen else 0, formatoceldaleft)
                        ws.write('AC%s' % filas_recorridas, f'{cincoacien(resumen.promedio_docencia_hetero)} %' if resumen else 0, formatoceldaleft)
                        ws.write('AD%s' % filas_recorridas, resumen.promedio_docencia_auto if resumen else 0, formatoceldaleft)
                        ws.write('AE%s' % filas_recorridas, f'{cincoacien(resumen.promedio_docencia_auto)} %' if resumen else 0, formatoceldaleft)
                        ws.write('AF%s' % filas_recorridas, resumen.promedio_docencia_directivo if resumen else 0, formatoceldaleft)
                        ws.write('AG%s' % filas_recorridas, f'{cincoacien(resumen.promedio_docencia_directivo)} %' if resumen else 0, formatoceldaleft)
                        ws.write('AH%s' % filas_recorridas, resumen.valor_tabla_docencia_hetero if resumen else 0, formatoceldaleft)
                        ws.write('AI%s' % filas_recorridas, f'{cincoacien(resumen.valor_tabla_docencia_hetero)} %' if resumen else 0, formatoceldaleft)
                        ws.write('AJ%s' % filas_recorridas, resumen.valor_tabla_docencia_auto if resumen else 0, formatoceldaleft)
                        ws.write('AK%s' % filas_recorridas, f'{cincoacien(resumen.valor_tabla_docencia_auto)} %' if resumen else 0, formatoceldaleft)
                        ws.write('AL%s' % filas_recorridas, resumen.valor_tabla_docencia_directivo if resumen else 0, formatoceldaleft)
                        ws.write('AM%s' % filas_recorridas, f'{cincoacien(resumen.valor_tabla_docencia_directivo)} %' if resumen else 0, formatoceldaleft)
                        ws.write('AN%s' % filas_recorridas, resumen.resultado_docencia if resumen else 0, formatoceldaleft)
                        ws.write('AO%s' % filas_recorridas, f'{cincoacien(resumen.resultado_docencia)} %' if resumen else 0, formato)

                        filas_recorridas += 1
                        print(f'{cont}/{eProfesorMaterias.count()}')
                        cont += 1

                    workbook.close()
                    output.seek(0)
                    filename = f'General_{datetime.now().date()}_{datetime.now().time()}.xlsx'
                    response = HttpResponse(output,

                                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    response['Content-Disposition'] = 'attachment; filename=%s' % filename
                    return response
                except Exception as ex:
                    pass

            elif action == 'descargarreportegráfica_sal':
                try:
                    __author__ = 'Unemi'

                    output = io.BytesIO()
                    workbook = xlsxwriter.Workbook(output)
                    ws = workbook.add_worksheet('salud')
                    ws.set_column(0, 0, 10)
                    ws.set_column(1, 1, 15)
                    ws.set_column(2, 5, 30)
                    ws.set_column(6, 6, 20)
                    ws.set_column(7, 7, 40)
                    ws.set_column(8, 9, 20)
                    ws.set_column(10, 40, 30)

                    formatotitulo_filtros = workbook.add_format(
                        {'bold': 1, 'text_wrap': True, 'border': 1, 'align': 'center', 'font_size': 14})

                    formatoceldacab = workbook.add_format(
                        {'align': 'center', 'bold': 1, 'border': 1, 'text_wrap': True, 'fg_color': '#1C3247', 'font_color': 'white'})

                    formatoceldaleft = workbook.add_format(
                        {'text_wrap': True, 'align': 'center', 'valign': 'vcenter', 'border': 1})

                    formatoceldaleft_red = workbook.add_format(
                        {'text_wrap': True, 'align': 'center', 'valign': 'vcenter', 'border': 1, 'fg_color': '#FF0000', 'font_color': 'black'})

                    formatoceldaleft_yelow = workbook.add_format(
                        {'text_wrap': True, 'align': 'center', 'valign': 'vcenter', 'border': 1, 'fg_color': '#FFFF00', 'font_color': 'black'})

                    formatoceldaleft_green = workbook.add_format(
                        {'text_wrap': True, 'align': 'center', 'valign': 'vcenter', 'border': 1, 'fg_color': '#92D050', 'font_color': 'black'})

                    desde = request.GET.get('desde', '')
                    hasta = request.GET.get('hasta', '')

                    carrera = None
                    if 'carrera' in request.GET:
                        carrera = int(request.GET['carrera'])

                    ws.merge_range('A1:AO1', 'SEGUIMIENTO DE ESTADÍSTICAS', formatotitulo_filtros)

                    ws.write(1, 0, 'N°', formatoceldacab)
                    ws.write(1, 1, 'ID', formatoceldacab)
                    ws.write(1, 2, 'Escuela', formatoceldacab)
                    ws.write(1, 3, 'Maestría', formatoceldacab)
                    ws.write(1, 4, 'Cohorte', formatoceldacab)
                    ws.write(1, 5, 'Módulo', formatoceldacab)
                    ws.write(1, 6, 'Paralelo', formatoceldacab)
                    ws.write(1, 7, 'Docente', formatoceldacab)
                    ws.write(1, 8, 'Inicio', formatoceldacab)
                    ws.write(1, 9, 'Fin', formatoceldacab)
                    ws.write(1, 10, 'Inicio eval_hetero', formatoceldacab)
                    ws.write(1, 11, 'Fin eval_hetero', formatoceldacab)
                    ws.write(1, 12, 'Estado eval_hetero', formatoceldacab)
                    ws.write(1, 13, 'Total evaluados', formatoceldacab)
                    ws.write(1, 14, 'Total matriculados', formatoceldacab)

                    ws.write(1, 15, 'Inicio eval_auto', formatoceldacab)
                    ws.write(1, 16, 'Fin eval_auto', formatoceldacab)
                    ws.write(1, 17, 'Estado eval_auto', formatoceldacab)
                    ws.write(1, 18, '¿Realizó Autoevaluación?', formatoceldacab)

                    ws.write(1, 19, 'Inicio eval_dir', formatoceldacab)
                    ws.write(1, 20, 'Fin eval_dir', formatoceldacab)
                    ws.write(1, 21, 'Estado eval_dir', formatoceldacab)
                    ws.write(1, 22, '¿Evaluado por Director?', formatoceldacab)
                    ws.write(1, 23, '¿Evaluado por Coordinador?', formatoceldacab)

                    ws.write(1, 24, '¿Procesado?', formatoceldacab)
                    ws.write(1, 25, 'Fecha de procesado', formatoceldacab)
                    ws.write(1, 26, 'Hora de procesado', formatoceldacab)

                    ws.write(1, 27, 'eval_hetero', formatoceldacab)
                    ws.write(1, 28, 'eval_hetero (%)', formatoceldacab)
                    ws.write(1, 29, 'eval_auto', formatoceldacab)
                    ws.write(1, 30, 'eval_auto (%)', formatoceldacab)
                    ws.write(1, 31, 'eval_directivo', formatoceldacab)
                    ws.write(1, 32, 'eval_directivo (%)', formatoceldacab)
                    ws.write(1, 33, 'evalp_hetero', formatoceldacab)
                    ws.write(1, 34, 'evalp_hetero (%)', formatoceldacab)
                    ws.write(1, 35, 'evalp_auto', formatoceldacab)
                    ws.write(1, 36, 'evalp_auto (%)', formatoceldacab)
                    ws.write(1, 37, 'evalp_directivo', formatoceldacab)
                    ws.write(1, 38, 'evalp_directivo (%)', formatoceldacab)
                    ws.write(1, 39, 'total_eval_abs', formatoceldacab)
                    ws.write(1, 40, 'total_eval_rel', formatoceldacab)

                    filtro = Q(status=True, tipoprofesor__id=11, materia__nivel__periodo__tipo__id=3, materia__fin__lt=datetime.now().date())

                    if carrera:
                        filtro = filtro & Q(materia__asignaturamalla__malla__carrera__id=carrera)

                    if desde and hasta:
                        filtro = filtro & Q(materia__fin__range=(desde, hasta))
                    elif desde:
                        filtro = filtro & Q(materia__fin__gte=desde)
                    elif hasta:
                        filtro = filtro & Q(materia__fin__lte=hasta)

                    eProfesorMateriasPrueba = ProfesorMateria.objects.filter(filtro).order_by('-id')
                    lis = []
                    cvnt = 0
                    for eModulo in eProfesorMateriasPrueba:
                        eProfesor = eModulo.profesor
                        distributivo = eProfesor.distributivohoraseval(eModulo.materia.nivel.periodo)
                        if distributivo and distributivo.resumen_evaluacion_acreditacion():
                            resumen = distributivo.resumen_evaluacion_acreditacion()
                            if resumen.resultado_docencia and resumen.resultado_docencia > 0:
                                if cincoacien(resumen.resultado_docencia) > 0 and cincoacien(resumen.resultado_docencia) <= 70:
                                    if eModulo.materia.asignaturamalla.malla.carrera.escuelaposgrado.id == 1:
                                        lis.append(eModulo.id)
                                        cvnt += 1
                                elif cincoacien(resumen.resultado_docencia) > 70 and cincoacien(resumen.resultado_docencia) <= 90:
                                    if eModulo.materia.asignaturamalla.malla.carrera.escuelaposgrado.id == 1:
                                        lis.append(eModulo.id)
                                        cvnt += 1
                                elif cincoacien(resumen.resultado_docencia) > 90:
                                    if eModulo.materia.asignaturamalla.malla.carrera.escuelaposgrado.id == 1:
                                        lis.append(eModulo.id)
                                        cvnt += 1
                                print(f'Contador: {cvnt}')

                    eProfesorMaterias = ProfesorMateria.objects.filter(id__in=lis).order_by('-id')
                    filas_recorridas = 3
                    cont = 1
                    for eProfesorMateria in eProfesorMaterias:
                        eMatriculados = eProfesorMateria.materia.materiaasignada_set.values("id").filter(status=True, matricula__estado_matricula__in=[2, 3], matricula__retiradomatricula=False).exclude(retiramateria=True).count()
                        eDir = object_dir(eProfesorMateria)
                        eRes = object_proces(eProfesorMateria)

                        distributivo = eProfesorMateria.profesor.distributivohoraseval(eProfesorMateria.materia.nivel.periodo)
                        if distributivo:
                            resumen = distributivo.resumen_evaluacion_acreditacion() if distributivo is not None else None
                        else:
                            resumen = None

                        formato = formatoceldaleft
                        if cincoacien(resumen.resultado_docencia) > 0 and cincoacien(resumen.resultado_docencia) <= 70:
                            formato = formatoceldaleft_red
                        elif cincoacien(resumen.resultado_docencia) > 70 and cincoacien(resumen.resultado_docencia) <= 90:
                            formato = formatoceldaleft_yelow
                        elif cincoacien(resumen.resultado_docencia) > 90:
                            formato = formatoceldaleft_green

                        ws.write('A%s' % filas_recorridas, str(cont), formatoceldaleft)
                        ws.write('B%s' % filas_recorridas, str(eProfesorMateria.id), formatoceldaleft)
                        ws.write('C%s' % filas_recorridas, str(eProfesorMateria.materia.asignaturamalla.malla.carrera.escuelaposgrado.nombre if eProfesorMateria.materia.asignaturamalla.malla.carrera.escuelaposgrado else 'NO REGISTRA'), formatoceldaleft)
                        ws.write('D%s' % filas_recorridas, str(eProfesorMateria.materia.asignaturamalla.malla.carrera), formatoceldaleft)
                        ws.write('E%s' % filas_recorridas, str(eProfesorMateria.materia.nivel.periodo.cohorte_maestria()), formatoceldaleft)
                        ws.write('F%s' % filas_recorridas, str(eProfesorMateria.materia.asignaturamalla.asignatura.nombre), formatoceldaleft)
                        ws.write('G%s' % filas_recorridas, str(eProfesorMateria.materia.paralelo), formatoceldaleft)
                        ws.write('H%s' % filas_recorridas, str(eProfesorMateria.profesor.persona), formatoceldaleft)
                        ws.write('I%s' % filas_recorridas, str(eProfesorMateria.materia.inicio), formatoceldaleft)
                        ws.write('J%s' % filas_recorridas, str(eProfesorMateria.materia.fin), formatoceldaleft)

                        ws.write('K%s' % filas_recorridas, str(eProfesorMateria.materia.inicioeval if eProfesorMateria.materia.inicioeval else 'NO CONFIGURADO'), formatoceldaleft)
                        ws.write('L%s' % filas_recorridas, str(eProfesorMateria.materia.fineval if eProfesorMateria.materia.fineval else 'NO CONFIGURADO'), formatoceldaleft)
                        ws.write('M%s' % filas_recorridas, str(fecha_vencida(eProfesorMateria.materia, 1)), formatoceldaleft)
                        ws.write('N%s' % filas_recorridas, cantidad_evaluacion_docente(eProfesorMateria), formatoceldaleft)
                        ws.write('O%s' % filas_recorridas, eMatriculados, formatoceldaleft)

                        ws.write('P%s' % filas_recorridas, str(eProfesorMateria.materia.inicioevalauto if eProfesorMateria.materia.inicioevalauto else 'NO CONFIGURADO'), formatoceldaleft)
                        ws.write('Q%s' % filas_recorridas, str(eProfesorMateria.materia.finevalauto if eProfesorMateria.materia.finevalauto else 'NO CONFIGURADO'), formatoceldaleft)
                        ws.write('R%s' % filas_recorridas, str(fecha_vencida(eProfesorMateria.materia, 2)), formatoceldaleft)
                        ws.write('S%s' % filas_recorridas, str('SI' if cantidad_evaluacion_auto(eProfesorMateria) > 0 else 'NO'), formatoceldaleft)

                        ws.write('T%s' % filas_recorridas, str(eDir.inicio if eDir else 'NO CONFIGURADO'), formatoceldaleft)
                        ws.write('U%s' % filas_recorridas, str(eDir.fin if eDir else 'NO CONFIGURADO'), formatoceldaleft)
                        ws.write('V%s' % filas_recorridas, str(fecha_vencida(eProfesorMateria.materia, 3)), formatoceldaleft)
                        ws.write('W%s' % filas_recorridas, str('SI' if evaluo_director(eProfesorMateria) else 'NO'), formatoceldaleft)
                        ws.write('X%s' % filas_recorridas, str('SI' if evaluo_coordinador(eProfesorMateria) else 'NO'), formatoceldaleft)

                        ws.write('Y%s' % filas_recorridas, str('SI' if eRes else 'NO PROCESADO'), formatoceldaleft)
                        ws.write('Z%s' % filas_recorridas, str(eRes.fecha_creacion.date().strftime('%d-%m-%Y') if eRes else 'NO PROCESADO'), formatoceldaleft)
                        ws.write('AA%s' % filas_recorridas, str(eRes.fecha_creacion.time().strftime('%I:%M %p') if eRes else 'NO PROCESADO'), formatoceldaleft)

                        ws.write('AB%s' % filas_recorridas, resumen.promedio_docencia_hetero if resumen else 0, formatoceldaleft)
                        ws.write('AC%s' % filas_recorridas, f'{cincoacien(resumen.promedio_docencia_hetero)} %' if resumen else 0, formatoceldaleft)
                        ws.write('AD%s' % filas_recorridas, resumen.promedio_docencia_auto if resumen else 0, formatoceldaleft)
                        ws.write('AE%s' % filas_recorridas, f'{cincoacien(resumen.promedio_docencia_auto)} %' if resumen else 0, formatoceldaleft)
                        ws.write('AF%s' % filas_recorridas, resumen.promedio_docencia_directivo if resumen else 0, formatoceldaleft)
                        ws.write('AG%s' % filas_recorridas, f'{cincoacien(resumen.promedio_docencia_directivo)} %' if resumen else 0, formatoceldaleft)
                        ws.write('AH%s' % filas_recorridas, resumen.valor_tabla_docencia_hetero if resumen else 0, formatoceldaleft)
                        ws.write('AI%s' % filas_recorridas, f'{cincoacien(resumen.valor_tabla_docencia_hetero)} %' if resumen else 0, formatoceldaleft)
                        ws.write('AJ%s' % filas_recorridas, resumen.valor_tabla_docencia_auto if resumen else 0, formatoceldaleft)
                        ws.write('AK%s' % filas_recorridas, f'{cincoacien(resumen.valor_tabla_docencia_auto)} %' if resumen else 0, formatoceldaleft)
                        ws.write('AL%s' % filas_recorridas, resumen.valor_tabla_docencia_directivo if resumen else 0, formatoceldaleft)
                        ws.write('AM%s' % filas_recorridas, f'{cincoacien(resumen.valor_tabla_docencia_directivo)} %' if resumen else 0, formatoceldaleft)
                        ws.write('AN%s' % filas_recorridas, resumen.resultado_docencia if resumen else 0, formatoceldaleft)
                        ws.write('AO%s' % filas_recorridas, f'{cincoacien(resumen.resultado_docencia)} %' if resumen else 0, formato)

                        filas_recorridas += 1
                        print(f'{cont}/{eProfesorMaterias.count()}')
                        cont += 1

                    workbook.close()
                    output.seek(0)
                    filename = f'Salud_{datetime.now().date()}_{datetime.now().time()}.xlsx'
                    response = HttpResponse(output,

                                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    response['Content-Disposition'] = 'attachment; filename=%s' % filename
                    return response
                except Exception as ex:
                    pass

            elif action == 'descargarreportegráfica_edu':
                try:
                    __author__ = 'Unemi'

                    output = io.BytesIO()
                    workbook = xlsxwriter.Workbook(output)
                    ws = workbook.add_worksheet('educacion')
                    ws.set_column(0, 0, 10)
                    ws.set_column(1, 1, 15)
                    ws.set_column(2, 5, 30)
                    ws.set_column(6, 6, 20)
                    ws.set_column(7, 7, 40)
                    ws.set_column(8, 9, 20)
                    ws.set_column(10, 40, 30)

                    formatotitulo_filtros = workbook.add_format(
                        {'bold': 1, 'text_wrap': True, 'border': 1, 'align': 'center', 'font_size': 14})

                    formatoceldacab = workbook.add_format(
                        {'align': 'center', 'bold': 1, 'border': 1, 'text_wrap': True, 'fg_color': '#1C3247', 'font_color': 'white'})

                    formatoceldaleft = workbook.add_format(
                        {'text_wrap': True, 'align': 'center', 'valign': 'vcenter', 'border': 1})

                    formatoceldaleft_red = workbook.add_format(
                        {'text_wrap': True, 'align': 'center', 'valign': 'vcenter', 'border': 1, 'fg_color': '#FF0000', 'font_color': 'black'})

                    formatoceldaleft_yelow = workbook.add_format(
                        {'text_wrap': True, 'align': 'center', 'valign': 'vcenter', 'border': 1, 'fg_color': '#FFFF00', 'font_color': 'black'})

                    formatoceldaleft_green = workbook.add_format(
                        {'text_wrap': True, 'align': 'center', 'valign': 'vcenter', 'border': 1, 'fg_color': '#92D050', 'font_color': 'black'})

                    desde = request.GET.get('desde', '')
                    hasta = request.GET.get('hasta', '')

                    carrera = None
                    if 'carrera' in request.GET:
                        carrera = int(request.GET['carrera'])

                    ws.merge_range('A1:AO1', 'SEGUIMIENTO DE ESTADÍSTICAS', formatotitulo_filtros)

                    ws.write(1, 0, 'N°', formatoceldacab)
                    ws.write(1, 1, 'ID', formatoceldacab)
                    ws.write(1, 2, 'Escuela', formatoceldacab)
                    ws.write(1, 3, 'Maestría', formatoceldacab)
                    ws.write(1, 4, 'Cohorte', formatoceldacab)
                    ws.write(1, 5, 'Módulo', formatoceldacab)
                    ws.write(1, 6, 'Paralelo', formatoceldacab)
                    ws.write(1, 7, 'Docente', formatoceldacab)
                    ws.write(1, 8, 'Inicio', formatoceldacab)
                    ws.write(1, 9, 'Fin', formatoceldacab)
                    ws.write(1, 10, 'Inicio eval_hetero', formatoceldacab)
                    ws.write(1, 11, 'Fin eval_hetero', formatoceldacab)
                    ws.write(1, 12, 'Estado eval_hetero', formatoceldacab)
                    ws.write(1, 13, 'Total evaluados', formatoceldacab)
                    ws.write(1, 14, 'Total matriculados', formatoceldacab)

                    ws.write(1, 15, 'Inicio eval_auto', formatoceldacab)
                    ws.write(1, 16, 'Fin eval_auto', formatoceldacab)
                    ws.write(1, 17, 'Estado eval_auto', formatoceldacab)
                    ws.write(1, 18, '¿Realizó Autoevaluación?', formatoceldacab)

                    ws.write(1, 19, 'Inicio eval_dir', formatoceldacab)
                    ws.write(1, 20, 'Fin eval_dir', formatoceldacab)
                    ws.write(1, 21, 'Estado eval_dir', formatoceldacab)
                    ws.write(1, 22, '¿Evaluado por Director?', formatoceldacab)
                    ws.write(1, 23, '¿Evaluado por Coordinador?', formatoceldacab)

                    ws.write(1, 24, '¿Procesado?', formatoceldacab)
                    ws.write(1, 25, 'Fecha de procesado', formatoceldacab)
                    ws.write(1, 26, 'Hora de procesado', formatoceldacab)

                    ws.write(1, 27, 'eval_hetero', formatoceldacab)
                    ws.write(1, 28, 'eval_hetero (%)', formatoceldacab)
                    ws.write(1, 29, 'eval_auto', formatoceldacab)
                    ws.write(1, 30, 'eval_auto (%)', formatoceldacab)
                    ws.write(1, 31, 'eval_directivo', formatoceldacab)
                    ws.write(1, 32, 'eval_directivo (%)', formatoceldacab)
                    ws.write(1, 33, 'evalp_hetero', formatoceldacab)
                    ws.write(1, 34, 'evalp_hetero (%)', formatoceldacab)
                    ws.write(1, 35, 'evalp_auto', formatoceldacab)
                    ws.write(1, 36, 'evalp_auto (%)', formatoceldacab)
                    ws.write(1, 37, 'evalp_directivo', formatoceldacab)
                    ws.write(1, 38, 'evalp_directivo (%)', formatoceldacab)
                    ws.write(1, 39, 'total_eval_abs', formatoceldacab)
                    ws.write(1, 40, 'total_eval_rel', formatoceldacab)

                    filtro = Q(status=True, tipoprofesor__id=11, materia__nivel__periodo__tipo__id=3, materia__fin__lt=datetime.now().date())

                    if carrera:
                        filtro = filtro & Q(materia__asignaturamalla__malla__carrera__id=carrera)

                    if desde and hasta:
                        filtro = filtro & Q(materia__fin__range=(desde, hasta))
                    elif desde:
                        filtro = filtro & Q(materia__fin__gte=desde)
                    elif hasta:
                        filtro = filtro & Q(materia__fin__lte=hasta)

                    eProfesorMateriasPrueba = ProfesorMateria.objects.filter(filtro).order_by('-id')
                    lis = []
                    cvnt = 0
                    for eModulo in eProfesorMateriasPrueba:
                        eProfesor = eModulo.profesor
                        distributivo = eProfesor.distributivohoraseval(eModulo.materia.nivel.periodo)
                        if distributivo and distributivo.resumen_evaluacion_acreditacion():
                            resumen = distributivo.resumen_evaluacion_acreditacion()
                            if resumen.resultado_docencia and resumen.resultado_docencia > 0:
                                if cincoacien(resumen.resultado_docencia) > 0 and cincoacien(resumen.resultado_docencia) <= 70:
                                    if eModulo.materia.asignaturamalla.malla.carrera.escuelaposgrado.id == 2:
                                        lis.append(eModulo.id)
                                        cvnt += 1
                                elif cincoacien(resumen.resultado_docencia) > 70 and cincoacien(resumen.resultado_docencia) <= 90:
                                    if eModulo.materia.asignaturamalla.malla.carrera.escuelaposgrado.id == 2:
                                        lis.append(eModulo.id)
                                        cvnt += 1
                                elif cincoacien(resumen.resultado_docencia) > 90:
                                    if eModulo.materia.asignaturamalla.malla.carrera.escuelaposgrado.id == 2:
                                        lis.append(eModulo.id)
                                        cvnt += 1
                                print(f'Contador: {cvnt}')

                    eProfesorMaterias = ProfesorMateria.objects.filter(id__in=lis).order_by('-id')
                    filas_recorridas = 3
                    cont = 1
                    for eProfesorMateria in eProfesorMaterias:
                        eMatriculados = eProfesorMateria.materia.materiaasignada_set.values("id").filter(status=True, matricula__estado_matricula__in=[2, 3], matricula__retiradomatricula=False).exclude(retiramateria=True).count()
                        eDir = object_dir(eProfesorMateria)
                        eRes = object_proces(eProfesorMateria)

                        distributivo = eProfesorMateria.profesor.distributivohoraseval(eProfesorMateria.materia.nivel.periodo)
                        if distributivo:
                            resumen = distributivo.resumen_evaluacion_acreditacion() if distributivo is not None else None
                        else:
                            resumen = None

                        formato = formatoceldaleft
                        if cincoacien(resumen.resultado_docencia) > 0 and cincoacien(resumen.resultado_docencia) <= 70:
                            formato = formatoceldaleft_red
                        elif cincoacien(resumen.resultado_docencia) > 70 and cincoacien(resumen.resultado_docencia) <= 90:
                            formato = formatoceldaleft_yelow
                        elif cincoacien(resumen.resultado_docencia) > 90:
                            formato = formatoceldaleft_green

                        ws.write('A%s' % filas_recorridas, str(cont), formatoceldaleft)
                        ws.write('B%s' % filas_recorridas, str(eProfesorMateria.id), formatoceldaleft)
                        ws.write('C%s' % filas_recorridas, str(eProfesorMateria.materia.asignaturamalla.malla.carrera.escuelaposgrado.nombre if eProfesorMateria.materia.asignaturamalla.malla.carrera.escuelaposgrado else 'NO REGISTRA'), formatoceldaleft)
                        ws.write('D%s' % filas_recorridas, str(eProfesorMateria.materia.asignaturamalla.malla.carrera), formatoceldaleft)
                        ws.write('E%s' % filas_recorridas, str(eProfesorMateria.materia.nivel.periodo.cohorte_maestria()), formatoceldaleft)
                        ws.write('F%s' % filas_recorridas, str(eProfesorMateria.materia.asignaturamalla.asignatura.nombre), formatoceldaleft)
                        ws.write('G%s' % filas_recorridas, str(eProfesorMateria.materia.paralelo), formatoceldaleft)
                        ws.write('H%s' % filas_recorridas, str(eProfesorMateria.profesor.persona), formatoceldaleft)
                        ws.write('I%s' % filas_recorridas, str(eProfesorMateria.materia.inicio), formatoceldaleft)
                        ws.write('J%s' % filas_recorridas, str(eProfesorMateria.materia.fin), formatoceldaleft)

                        ws.write('K%s' % filas_recorridas, str(eProfesorMateria.materia.inicioeval if eProfesorMateria.materia.inicioeval else 'NO CONFIGURADO'), formatoceldaleft)
                        ws.write('L%s' % filas_recorridas, str(eProfesorMateria.materia.fineval if eProfesorMateria.materia.fineval else 'NO CONFIGURADO'), formatoceldaleft)
                        ws.write('M%s' % filas_recorridas, str(fecha_vencida(eProfesorMateria.materia, 1)), formatoceldaleft)
                        ws.write('N%s' % filas_recorridas, cantidad_evaluacion_docente(eProfesorMateria), formatoceldaleft)
                        ws.write('O%s' % filas_recorridas, eMatriculados, formatoceldaleft)

                        ws.write('P%s' % filas_recorridas, str(eProfesorMateria.materia.inicioevalauto if eProfesorMateria.materia.inicioevalauto else 'NO CONFIGURADO'), formatoceldaleft)
                        ws.write('Q%s' % filas_recorridas, str(eProfesorMateria.materia.finevalauto if eProfesorMateria.materia.finevalauto else 'NO CONFIGURADO'), formatoceldaleft)
                        ws.write('R%s' % filas_recorridas, str(fecha_vencida(eProfesorMateria.materia, 2)), formatoceldaleft)
                        ws.write('S%s' % filas_recorridas, str('SI' if cantidad_evaluacion_auto(eProfesorMateria) > 0 else 'NO'), formatoceldaleft)

                        ws.write('T%s' % filas_recorridas, str(eDir.inicio if eDir else 'NO CONFIGURADO'), formatoceldaleft)
                        ws.write('U%s' % filas_recorridas, str(eDir.fin if eDir else 'NO CONFIGURADO'), formatoceldaleft)
                        ws.write('V%s' % filas_recorridas, str(fecha_vencida(eProfesorMateria.materia, 3)), formatoceldaleft)
                        ws.write('W%s' % filas_recorridas, str('SI' if evaluo_director(eProfesorMateria) else 'NO'), formatoceldaleft)
                        ws.write('X%s' % filas_recorridas, str('SI' if evaluo_coordinador(eProfesorMateria) else 'NO'), formatoceldaleft)

                        ws.write('Y%s' % filas_recorridas, str('SI' if eRes else 'NO PROCESADO'), formatoceldaleft)
                        ws.write('Z%s' % filas_recorridas, str(eRes.fecha_creacion.date().strftime('%d-%m-%Y') if eRes else 'NO PROCESADO'), formatoceldaleft)
                        ws.write('AA%s' % filas_recorridas, str(eRes.fecha_creacion.time().strftime('%I:%M %p') if eRes else 'NO PROCESADO'), formatoceldaleft)

                        ws.write('AB%s' % filas_recorridas, resumen.promedio_docencia_hetero if resumen else 0, formatoceldaleft)
                        ws.write('AC%s' % filas_recorridas, f'{cincoacien(resumen.promedio_docencia_hetero)} %' if resumen else 0, formatoceldaleft)
                        ws.write('AD%s' % filas_recorridas, resumen.promedio_docencia_auto if resumen else 0, formatoceldaleft)
                        ws.write('AE%s' % filas_recorridas, f'{cincoacien(resumen.promedio_docencia_auto)} %' if resumen else 0, formatoceldaleft)
                        ws.write('AF%s' % filas_recorridas, resumen.promedio_docencia_directivo if resumen else 0, formatoceldaleft)
                        ws.write('AG%s' % filas_recorridas, f'{cincoacien(resumen.promedio_docencia_directivo)} %' if resumen else 0, formatoceldaleft)
                        ws.write('AH%s' % filas_recorridas, resumen.valor_tabla_docencia_hetero if resumen else 0, formatoceldaleft)
                        ws.write('AI%s' % filas_recorridas, f'{cincoacien(resumen.valor_tabla_docencia_hetero)} %' if resumen else 0, formatoceldaleft)
                        ws.write('AJ%s' % filas_recorridas, resumen.valor_tabla_docencia_auto if resumen else 0, formatoceldaleft)
                        ws.write('AK%s' % filas_recorridas, f'{cincoacien(resumen.valor_tabla_docencia_auto)} %' if resumen else 0, formatoceldaleft)
                        ws.write('AL%s' % filas_recorridas, resumen.valor_tabla_docencia_directivo if resumen else 0, formatoceldaleft)
                        ws.write('AM%s' % filas_recorridas, f'{cincoacien(resumen.valor_tabla_docencia_directivo)} %' if resumen else 0, formatoceldaleft)
                        ws.write('AN%s' % filas_recorridas, resumen.resultado_docencia if resumen else 0, formatoceldaleft)
                        ws.write('AO%s' % filas_recorridas, f'{cincoacien(resumen.resultado_docencia)} %' if resumen else 0, formato)

                        filas_recorridas += 1
                        print(f'{cont}/{eProfesorMaterias.count()}')
                        cont += 1

                    workbook.close()
                    output.seek(0)
                    filename = f'Educacion_{datetime.now().date()}_{datetime.now().time()}.xlsx'
                    response = HttpResponse(output,

                                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    response['Content-Disposition'] = 'attachment; filename=%s' % filename
                    return response
                except Exception as ex:
                    pass

            elif action == 'descargarreportegráfica_neg':
                try:
                    __author__ = 'Unemi'

                    output = io.BytesIO()
                    workbook = xlsxwriter.Workbook(output)
                    ws = workbook.add_worksheet('negocios')
                    ws.set_column(0, 0, 10)
                    ws.set_column(1, 1, 15)
                    ws.set_column(2, 5, 30)
                    ws.set_column(6, 6, 20)
                    ws.set_column(7, 7, 40)
                    ws.set_column(8, 9, 20)
                    ws.set_column(10, 40, 30)

                    formatotitulo_filtros = workbook.add_format(
                        {'bold': 1, 'text_wrap': True, 'border': 1, 'align': 'center', 'font_size': 14})

                    formatoceldacab = workbook.add_format(
                        {'align': 'center', 'bold': 1, 'border': 1, 'text_wrap': True, 'fg_color': '#1C3247', 'font_color': 'white'})

                    formatoceldaleft = workbook.add_format(
                        {'text_wrap': True, 'align': 'center', 'valign': 'vcenter', 'border': 1})

                    formatoceldaleft_red = workbook.add_format(
                        {'text_wrap': True, 'align': 'center', 'valign': 'vcenter', 'border': 1, 'fg_color': '#FF0000', 'font_color': 'black'})

                    formatoceldaleft_yelow = workbook.add_format(
                        {'text_wrap': True, 'align': 'center', 'valign': 'vcenter', 'border': 1, 'fg_color': '#FFFF00', 'font_color': 'black'})

                    formatoceldaleft_green = workbook.add_format(
                        {'text_wrap': True, 'align': 'center', 'valign': 'vcenter', 'border': 1, 'fg_color': '#92D050', 'font_color': 'black'})

                    desde = request.GET.get('desde', '')
                    hasta = request.GET.get('hasta', '')

                    carrera = None
                    if 'carrera' in request.GET:
                        carrera = int(request.GET['carrera'])

                    ws.merge_range('A1:AO1', 'SEGUIMIENTO DE ESTADÍSTICAS', formatotitulo_filtros)

                    ws.write(1, 0, 'N°', formatoceldacab)
                    ws.write(1, 1, 'ID', formatoceldacab)
                    ws.write(1, 2, 'Escuela', formatoceldacab)
                    ws.write(1, 3, 'Maestría', formatoceldacab)
                    ws.write(1, 4, 'Cohorte', formatoceldacab)
                    ws.write(1, 5, 'Módulo', formatoceldacab)
                    ws.write(1, 6, 'Paralelo', formatoceldacab)
                    ws.write(1, 7, 'Docente', formatoceldacab)
                    ws.write(1, 8, 'Inicio', formatoceldacab)
                    ws.write(1, 9, 'Fin', formatoceldacab)
                    ws.write(1, 10, 'Inicio eval_hetero', formatoceldacab)
                    ws.write(1, 11, 'Fin eval_hetero', formatoceldacab)
                    ws.write(1, 12, 'Estado eval_hetero', formatoceldacab)
                    ws.write(1, 13, 'Total evaluados', formatoceldacab)
                    ws.write(1, 14, 'Total matriculados', formatoceldacab)

                    ws.write(1, 15, 'Inicio eval_auto', formatoceldacab)
                    ws.write(1, 16, 'Fin eval_auto', formatoceldacab)
                    ws.write(1, 17, 'Estado eval_auto', formatoceldacab)
                    ws.write(1, 18, '¿Realizó Autoevaluación?', formatoceldacab)

                    ws.write(1, 19, 'Inicio eval_dir', formatoceldacab)
                    ws.write(1, 20, 'Fin eval_dir', formatoceldacab)
                    ws.write(1, 21, 'Estado eval_dir', formatoceldacab)
                    ws.write(1, 22, '¿Evaluado por Director?', formatoceldacab)
                    ws.write(1, 23, '¿Evaluado por Coordinador?', formatoceldacab)

                    ws.write(1, 24, '¿Procesado?', formatoceldacab)
                    ws.write(1, 25, 'Fecha de procesado', formatoceldacab)
                    ws.write(1, 26, 'Hora de procesado', formatoceldacab)

                    ws.write(1, 27, 'eval_hetero', formatoceldacab)
                    ws.write(1, 28, 'eval_hetero (%)', formatoceldacab)
                    ws.write(1, 29, 'eval_auto', formatoceldacab)
                    ws.write(1, 30, 'eval_auto (%)', formatoceldacab)
                    ws.write(1, 31, 'eval_directivo', formatoceldacab)
                    ws.write(1, 32, 'eval_directivo (%)', formatoceldacab)
                    ws.write(1, 33, 'evalp_hetero', formatoceldacab)
                    ws.write(1, 34, 'evalp_hetero (%)', formatoceldacab)
                    ws.write(1, 35, 'evalp_auto', formatoceldacab)
                    ws.write(1, 36, 'evalp_auto (%)', formatoceldacab)
                    ws.write(1, 37, 'evalp_directivo', formatoceldacab)
                    ws.write(1, 38, 'evalp_directivo (%)', formatoceldacab)
                    ws.write(1, 39, 'total_eval_abs', formatoceldacab)
                    ws.write(1, 40, 'total_eval_rel', formatoceldacab)

                    filtro = Q(status=True, tipoprofesor__id=11, materia__nivel__periodo__tipo__id=3, materia__fin__lt=datetime.now().date())

                    if carrera:
                        filtro = filtro & Q(materia__asignaturamalla__malla__carrera__id=carrera)

                    if desde and hasta:
                        filtro = filtro & Q(materia__fin__range=(desde, hasta))
                    elif desde:
                        filtro = filtro & Q(materia__fin__gte=desde)
                    elif hasta:
                        filtro = filtro & Q(materia__fin__lte=hasta)

                    eProfesorMateriasPrueba = ProfesorMateria.objects.filter(filtro).order_by('-id')
                    lis = []
                    cvnt = 0
                    for eModulo in eProfesorMateriasPrueba:
                        eProfesor = eModulo.profesor
                        distributivo = eProfesor.distributivohoraseval(eModulo.materia.nivel.periodo)
                        if distributivo and distributivo.resumen_evaluacion_acreditacion():
                            resumen = distributivo.resumen_evaluacion_acreditacion()
                            if resumen.resultado_docencia and resumen.resultado_docencia > 0:
                                if cincoacien(resumen.resultado_docencia) > 0 and cincoacien(resumen.resultado_docencia) <= 70:
                                    if eModulo.materia.asignaturamalla.malla.carrera.escuelaposgrado.id == 3:
                                        lis.append(eModulo.id)
                                        cvnt += 1
                                elif cincoacien(resumen.resultado_docencia) > 70 and cincoacien(resumen.resultado_docencia) <= 90:
                                    if eModulo.materia.asignaturamalla.malla.carrera.escuelaposgrado.id == 3:
                                        lis.append(eModulo.id)
                                        cvnt += 1
                                elif cincoacien(resumen.resultado_docencia) > 90:
                                    if eModulo.materia.asignaturamalla.malla.carrera.escuelaposgrado.id == 3:
                                        lis.append(eModulo.id)
                                        cvnt += 1
                                print(f'Contador: {cvnt}')

                    eProfesorMaterias = ProfesorMateria.objects.filter(id__in=lis).order_by('-id')
                    filas_recorridas = 3
                    cont = 1
                    for eProfesorMateria in eProfesorMaterias:
                        eMatriculados = eProfesorMateria.materia.materiaasignada_set.values("id").filter(status=True, matricula__estado_matricula__in=[2, 3], matricula__retiradomatricula=False).exclude(retiramateria=True).count()
                        eDir = object_dir(eProfesorMateria)
                        eRes = object_proces(eProfesorMateria)

                        distributivo = eProfesorMateria.profesor.distributivohoraseval(eProfesorMateria.materia.nivel.periodo)
                        if distributivo:
                            resumen = distributivo.resumen_evaluacion_acreditacion() if distributivo is not None else None
                        else:
                            resumen = None

                        formato = formatoceldaleft
                        if cincoacien(resumen.resultado_docencia) > 0 and cincoacien(resumen.resultado_docencia) <= 70:
                            formato = formatoceldaleft_red
                        elif cincoacien(resumen.resultado_docencia) > 70 and cincoacien(resumen.resultado_docencia) <= 90:
                            formato = formatoceldaleft_yelow
                        elif cincoacien(resumen.resultado_docencia) > 90:
                            formato = formatoceldaleft_green

                        ws.write('A%s' % filas_recorridas, str(cont), formatoceldaleft)
                        ws.write('B%s' % filas_recorridas, str(eProfesorMateria.id), formatoceldaleft)
                        ws.write('C%s' % filas_recorridas, str(eProfesorMateria.materia.asignaturamalla.malla.carrera.escuelaposgrado.nombre if eProfesorMateria.materia.asignaturamalla.malla.carrera.escuelaposgrado else 'NO REGISTRA'), formatoceldaleft)
                        ws.write('D%s' % filas_recorridas, str(eProfesorMateria.materia.asignaturamalla.malla.carrera), formatoceldaleft)
                        ws.write('E%s' % filas_recorridas, str(eProfesorMateria.materia.nivel.periodo.cohorte_maestria()), formatoceldaleft)
                        ws.write('F%s' % filas_recorridas, str(eProfesorMateria.materia.asignaturamalla.asignatura.nombre), formatoceldaleft)
                        ws.write('G%s' % filas_recorridas, str(eProfesorMateria.materia.paralelo), formatoceldaleft)
                        ws.write('H%s' % filas_recorridas, str(eProfesorMateria.profesor.persona), formatoceldaleft)
                        ws.write('I%s' % filas_recorridas, str(eProfesorMateria.materia.inicio), formatoceldaleft)
                        ws.write('J%s' % filas_recorridas, str(eProfesorMateria.materia.fin), formatoceldaleft)

                        ws.write('K%s' % filas_recorridas, str(eProfesorMateria.materia.inicioeval if eProfesorMateria.materia.inicioeval else 'NO CONFIGURADO'), formatoceldaleft)
                        ws.write('L%s' % filas_recorridas, str(eProfesorMateria.materia.fineval if eProfesorMateria.materia.fineval else 'NO CONFIGURADO'), formatoceldaleft)
                        ws.write('M%s' % filas_recorridas, str(fecha_vencida(eProfesorMateria.materia, 1)), formatoceldaleft)
                        ws.write('N%s' % filas_recorridas, cantidad_evaluacion_docente(eProfesorMateria), formatoceldaleft)
                        ws.write('O%s' % filas_recorridas, eMatriculados, formatoceldaleft)

                        ws.write('P%s' % filas_recorridas, str(eProfesorMateria.materia.inicioevalauto if eProfesorMateria.materia.inicioevalauto else 'NO CONFIGURADO'), formatoceldaleft)
                        ws.write('Q%s' % filas_recorridas, str(eProfesorMateria.materia.finevalauto if eProfesorMateria.materia.finevalauto else 'NO CONFIGURADO'), formatoceldaleft)
                        ws.write('R%s' % filas_recorridas, str(fecha_vencida(eProfesorMateria.materia, 2)), formatoceldaleft)
                        ws.write('S%s' % filas_recorridas, str('SI' if cantidad_evaluacion_auto(eProfesorMateria) > 0 else 'NO'), formatoceldaleft)

                        ws.write('T%s' % filas_recorridas, str(eDir.inicio if eDir else 'NO CONFIGURADO'), formatoceldaleft)
                        ws.write('U%s' % filas_recorridas, str(eDir.fin if eDir else 'NO CONFIGURADO'), formatoceldaleft)
                        ws.write('V%s' % filas_recorridas, str(fecha_vencida(eProfesorMateria.materia, 3)), formatoceldaleft)
                        ws.write('W%s' % filas_recorridas, str('SI' if evaluo_director(eProfesorMateria) else 'NO'), formatoceldaleft)
                        ws.write('X%s' % filas_recorridas, str('SI' if evaluo_coordinador(eProfesorMateria) else 'NO'), formatoceldaleft)

                        ws.write('Y%s' % filas_recorridas, str('SI' if eRes else 'NO PROCESADO'), formatoceldaleft)
                        ws.write('Z%s' % filas_recorridas, str(eRes.fecha_creacion.date().strftime('%d-%m-%Y') if eRes else 'NO PROCESADO'), formatoceldaleft)
                        ws.write('AA%s' % filas_recorridas, str(eRes.fecha_creacion.time().strftime('%I:%M %p') if eRes else 'NO PROCESADO'), formatoceldaleft)

                        ws.write('AB%s' % filas_recorridas, resumen.promedio_docencia_hetero if resumen else 0, formatoceldaleft)
                        ws.write('AC%s' % filas_recorridas, f'{cincoacien(resumen.promedio_docencia_hetero)} %' if resumen else 0, formatoceldaleft)
                        ws.write('AD%s' % filas_recorridas, resumen.promedio_docencia_auto if resumen else 0, formatoceldaleft)
                        ws.write('AE%s' % filas_recorridas, f'{cincoacien(resumen.promedio_docencia_auto)} %' if resumen else 0, formatoceldaleft)
                        ws.write('AF%s' % filas_recorridas, resumen.promedio_docencia_directivo if resumen else 0, formatoceldaleft)
                        ws.write('AG%s' % filas_recorridas, f'{cincoacien(resumen.promedio_docencia_directivo)} %' if resumen else 0, formatoceldaleft)
                        ws.write('AH%s' % filas_recorridas, resumen.valor_tabla_docencia_hetero if resumen else 0, formatoceldaleft)
                        ws.write('AI%s' % filas_recorridas, f'{cincoacien(resumen.valor_tabla_docencia_hetero)} %' if resumen else 0, formatoceldaleft)
                        ws.write('AJ%s' % filas_recorridas, resumen.valor_tabla_docencia_auto if resumen else 0, formatoceldaleft)
                        ws.write('AK%s' % filas_recorridas, f'{cincoacien(resumen.valor_tabla_docencia_auto)} %' if resumen else 0, formatoceldaleft)
                        ws.write('AL%s' % filas_recorridas, resumen.valor_tabla_docencia_directivo if resumen else 0, formatoceldaleft)
                        ws.write('AM%s' % filas_recorridas, f'{cincoacien(resumen.valor_tabla_docencia_directivo)} %' if resumen else 0, formatoceldaleft)
                        ws.write('AN%s' % filas_recorridas, resumen.resultado_docencia if resumen else 0, formatoceldaleft)
                        ws.write('AO%s' % filas_recorridas, f'{cincoacien(resumen.resultado_docencia)} %' if resumen else 0, formato)

                        filas_recorridas += 1
                        print(f'{cont}/{eProfesorMaterias.count()}')
                        cont += 1

                    workbook.close()
                    output.seek(0)
                    filename = f'Negocios_{datetime.now().date()}_{datetime.now().time()}.xlsx'
                    response = HttpResponse(output,

                                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    response['Content-Disposition'] = 'attachment; filename=%s' % filename
                    return response
                except Exception as ex:
                    pass

            elif action == 'mapeoleadscanales':
                try:
                    data['title'] = u'Mapeo de leads por canales de información'
                    listado_reportadas = []
                    listado_reportadas1 = []
                    listado_canales = []
                    tot = 0
                    tot1 = 0

                    maestria = cohorte = asesorcomercial = pais = provincia = canton = 0
                    desde = hasta = ''

                    filtro = Q(status=True)
                    filtro1 = Q(status=False)

                    if 'maestria' in request.GET:
                        maestria = int(request.GET['maestria'])
                    if 'cohorte' in request.GET:
                        cohorte = int(request.GET['cohorte'])
                    if 'asesor' in request.GET:
                        asesorcomercial = int(request.GET['asesor'])
                    if 'pais' in request.GET:
                        pais = int(request.GET['pais'])
                    if 'provincia' in request.GET:
                        provincia = int(request.GET['provincia'])
                    if 'canton' in request.GET:
                        canton = int(request.GET['canton'])
                    if 'desde' in request.GET:
                        desde = request.GET['desde']
                    if 'hasta' in request.GET:
                        hasta = request.GET['hasta']

                    url_vars = ' '
                    cohortes = provincias = cantones= None

                    if maestria > 0:
                        data['maestria'] = maestria
                        filtro = filtro & Q(cohortes__maestriaadmision__id=maestria)
                        filtro1 = filtro1 & Q(cohortes__maestriaadmision__id=maestria)
                        cohortes = CohorteMaestria.objects.filter(maestriaadmision__id=maestria)
                        url_vars += "&maestria={}".format(maestria)

                    if cohorte > 0:
                        data['cohorte'] = cohorte
                        filtro = filtro & Q(cohortes__id=cohorte)
                        filtro1 = filtro1 & Q(cohortes__id=cohorte)
                        url_vars += "&cohorte={}".format(cohorte)

                    if asesorcomercial > 0:
                        data['asesor'] = asesorcomercial
                        filtro = filtro & Q(asesor__id=asesorcomercial)
                        filtro1 = filtro1 & Q(asesor__id=asesorcomercial)
                        url_vars += "&asesor={}".format(asesorcomercial)

                    if pais > 0:
                        data['pais'] = pais
                        filtro = filtro & Q(inscripcionaspirante__persona__pais__id=pais)
                        filtro1 = filtro1 & Q(inscripcionaspirante__persona__pais__id=pais)
                        provincias = Provincia.objects.filter(pais__id=pais)
                        url_vars += "&pais={}".format(pais)

                    if provincia > 0:
                        data['provincia'] = provincia
                        filtro = filtro & Q(inscripcionaspirante__persona__provincia__id=provincia)
                        filtro1 = filtro1 & Q(inscripcionaspirante__persona__provincia__id=provincia)
                        cantones = Canton.objects.filter(provincia__id=provincia)
                        url_vars += "&provincia={}".format(provincia)

                    if canton > 0:
                        data['canton'] = canton
                        filtro = filtro & Q(inscripcionaspirante__persona__canton__id=canton)
                        filtro1 = filtro1 & Q(inscripcionaspirante__persona__canton__id=canton)
                        url_vars += "&canton={}".format(canton)

                    if desde and hasta:
                        data['desde'] = desde
                        data['hasta'] = hasta
                        url_vars += "&desde={}".format(desde)
                        url_vars += "&hasta={}".format(hasta)
                        filtro = filtro & Q(fecha_creacion__range=(desde, hasta))
                        filtro1 = filtro1 & Q(fecha_creacion__range=(desde, hasta))

                    elif desde:
                        data['desde'] = desde
                        url_vars += "&desde={}".format(desde)
                        filtro = filtro & Q(fecha_creacion__gte=desde)
                        filtro1 = filtro1 & Q(fecha_creacion__gte=desde)
                    elif hasta:
                        data['hasta'] = hasta
                        url_vars += "&hasta={}".format(hasta)
                        filtro = filtro & Q(fecha_creacion__lte=hasta)
                        filtro1 = filtro1 & Q(fecha_creacion__lte=hasta)

                    if maestria == 0 and cohorte == 0 and asesorcomercial == 0 and pais == 0 and provincia == 0 and canton == 0 and desde == '' and hasta == '':
                        leads = []
                        leads1 = []
                        canales = []
                    else:
                        leads = InscripcionCohorte.objects.filter(filtro).order_by('-id')
                        leads1 = InscripcionCohorte.objects.filter(filtro1).order_by('-id')
                        listado_reportadas.append(leads.values_list('id', flat=True))
                        listado_reportadas1.append(leads1.values_list('id', flat=True))
                        tot = leads.count()
                        tot1 = leads1.count()
                        canales = CanalInformacionMaestria.objects.filter(status=True, valido_form=True)
                    dicc = {}
                    for canal in canales:
                        qty2 = InscripcionCohorte.objects.filter(id__in=listado_reportadas, canal=canal).count()
                        qty3 = InscripcionCohorte.objects.filter(id__in=listado_reportadas1, canal=canal).count()
                        dicc = {'canal': canal.descripcion, 'cantidad': qty2, 'decan': qty3}
                        listado_canales.append(dicc)

                    if len(listado_reportadas) > 0:
                        qtnulo = InscripcionCohorte.objects.filter(id__in=listado_reportadas, canal__isnull=True).count()
                        qtnulo3 = InscripcionCohorte.objects.filter(id__in=listado_reportadas1, canal__isnull=True).count()
                        dicc = {'canal': 'Sin canal', 'cantidad': qtnulo, 'decan':qtnulo3}
                        listado_canales.append(dicc)

                    newlist = sorted(listado_canales, key=lambda d: d['cantidad'], reverse=True)
                    lis_canales = [d['canal'] for d in newlist]
                    lis_can = [d['cantidad'] for d in newlist]
                    lis_decan = [d['decan'] for d in newlist]

                    data['listado_reportadas'] = listado_reportadas
                    data['labelcanales'] = lis_canales
                    data['cohortesmaestria'] = CohorteMaestria.objects.filter(status=True)
                    data['cant'] = lis_can
                    data['decan'] = lis_decan
                    data['maestrialist'] = MaestriasAdmision.objects.filter(status=True, carrera__coordinacion__id=7)
                    data['cohorteslist'] = cohortes
                    data['asesorlist'] = AsesorComercial.objects.filter(status=True)
                    data['paislist'] = Pais.objects.filter(status=True)
                    data['provincialist'] = provincias
                    data['cantonlist'] = cantones
                    data['url_vars'] = url_vars
                    data['total_leads'] = tot + tot1
                    data['tot'] = tot
                    data['tot1'] = tot1
                    return render(request, "estadisticas/mapeoleadscanales.html", data)
                except Exception as ex:
                    pass

            elif action == 'matriculadosrangoedadescoor':
                try:
                    data['title'] = u'Matriculados por coordinaciones segun rango de edades'
                    data['coordinaciones'] = [c for c in Coordinacion.objects.all().order_by('id') if c.cantidad_matriculados(periodo)]
                    data['matriculados_menor_30'] = total_matriculados_menor_30(periodo)
                    data['matriculados_31_40'] = total_matriculados_31_40(periodo)
                    data['matriculados_41_50'] = total_matriculados_41_50(periodo)
                    data['matriculados_51_60'] = total_matriculados_51_60(periodo)
                    data['matriculados_mayor_61'] = total_matriculados_mayor_61(periodo)
                    data['total_matriculados'] = total_matriculados(periodo)
                    return render(request, "estadisticas/matriculadosrangoedadescoor.html", data)
                except Exception as ex:
                    pass

            elif action == 'matriculadosbecadiscapacidadcarr':
                try:
                    data['title'] = u'Matriculados por carreras con beca o discapacidad'
                    data['carreras'] = [c for c in Carrera.objects.filter(status=True).order_by('id') if c.cantidad_matriculados(periodo)]
                    data['cantidad_matriculados_beca'] = cantidad_matriculados_beca(periodo)
                    data['porciento_matriculados_beca'] = porciento_matriculados_beca(periodo)
                    data['cantidad_matriculados_discapacidad'] = cantidad_matriculados_discapacidad(periodo)
                    data['porciento_matriculados_discapacidad'] = porciento_matriculados_discapacidad(periodo)
                    data['total_matriculados'] = total_matriculados(periodo)
                    return render(request, "estadisticas/matriculadosbecadiscapacidadcarr.html", data)
                except Exception as ex:
                    pass

            elif action == 'matriculadosbecadiscapacidadcoor':
                try:
                    data['title'] = u'Matriculados por coordinaciones con beca o discapacidad'
                    data['coordinaciones'] = [c for c in Coordinacion.objects.filter(status=True).order_by('id') if c.cantidad_matriculados(periodo)]
                    data['cantidad_matriculados_beca'] = cantidad_matriculados_beca(periodo)
                    data['porciento_matriculados_beca'] = porciento_matriculados_beca(periodo)
                    data['cantidad_matriculados_discapacidad'] = cantidad_matriculados_discapacidad(periodo)
                    data['porciento_matriculados_discapacidad'] = porciento_matriculados_discapacidad(periodo)
                    data['total_matriculados'] = total_matriculados(periodo)
                    return render(request, "estadisticas/matriculadosbecadiscapacidadcoor.html", data)
                except Exception as ex:
                    pass

            elif action == 'matriculadosgeocarr':
                try:
                    data['title'] = u'Matriculados por carreras segun provincia y cantones'
                    data['carreras'] = [c for c in Carrera.objects.filter(status=True).order_by('id') if c.cantidad_matriculados(periodo)]
                    provincias = [x for x in Provincia.objects.filter(status=True) if x.cantidad_matriculados(periodo)]
                    data['provincias'] = provincias
                    cantones = [y for y in Canton.objects.filter(status=True) if y.cantidad_matriculados(periodo)]
                    data['cantones'] = cantones
                    listaxprovincia = []
                    for provincia in provincias:
                        fila = []
                        for carrera in data['carreras']:
                            fila.append(carrera.cantidad_matriculados_provincia(provincia, periodo))
                        listaxprovincia.append((provincia.nombre, fila, provincia.cantidad_matriculados(periodo), provincia.porciento_matriculados(periodo)))
                    listaxcanton = []
                    for canton in cantones:
                        fila = []
                        for carrera in data['carreras']:
                            fila.append(carrera.cantidad_matriculados_canton(canton, periodo))
                        listaxcanton.append((canton.nombre, fila, canton.cantidad_matriculados(periodo), canton.porciento_matriculados(periodo)))
                    data['listaxprovincia'] = listaxprovincia
                    data['listaxcanton'] = listaxcanton
                    data['total_matriculados'] = total_matriculados(periodo)
                    return render(request, "estadisticas/matriculadosgeocarr.html", data)
                except Exception as ex:
                    pass

            elif action == 'matriculadosgeocoor':
                try:
                    data['title'] = u'Matriculados por coordinaciones segun provincia y cantones'
                    data['coordinaciones'] = [c for c in Coordinacion.objects.filter(status=True).order_by('id') if c.cantidad_matriculados(periodo)]
                    provincias = [x for x in Provincia.objects.filter(status=True) if x.cantidad_matriculados(periodo)]
                    data['provincias'] = provincias
                    cantones = [y for y in Canton.objects.filter(status=True) if y.cantidad_matriculados(periodo)]
                    data['cantones'] = cantones
                    listaxprovincia = []
                    for provincia in provincias:
                        fila = []
                        for coordinacion in data['coordinaciones']:
                            fila.append(coordinacion.cantidad_matriculados_provincia(provincia, periodo))
                        listaxprovincia.append((provincia.nombre, fila, provincia.cantidad_matriculados(periodo), provincia.porciento_matriculados(periodo)))
                    listaxcanton = []
                    for canton in cantones:
                        fila = []
                        for coordinacion in data['coordinaciones']:
                            fila.append(coordinacion.cantidad_matriculados_canton(canton, periodo))
                        listaxcanton.append((canton.nombre, fila, canton.cantidad_matriculados(periodo), canton.porciento_matriculados(periodo)))
                    data['listaxprovincia'] = listaxprovincia
                    data['listaxcanton'] = listaxcanton
                    data['total_matriculados'] = total_matriculados(periodo)
                    return render(request, "estadisticas/matriculadosgeocoor.html", data)
                except Exception as ex:
                    pass

            elif action == 'reportexlsx':
                try:
                    filtro_carrera, filtro_desde, filtros, filtros_c, url = request.GET.get('car', ''), request.GET.get('desde', ''), Q(status=True), Q(status=True), ''
                    filtro_hasta = None
                    if request.GET.get('hasta', ''):
                        filtro_hasta = u"%s %s" % (request.GET.get('hasta', ''), '23:59:59')
                    if filtro_carrera:
                        data['carrera_filtrada'] = carrera_id = int(filtro_carrera)
                        url += "&carrera={}".format(filtro_carrera)
                        filtros = filtros & Q(cohortes__maestriaadmision__carrera_id=carrera_id)
                    # else:
                    #     filtros_c = filtros_c & Q(id)
                    if filtro_desde and filtro_hasta:
                        data['desde_filtrado'] = filtro_desde
                        data['hasta_filtrado'] = request.GET.get('hasta', '')
                        url += "&desde={}".format(filtro_desde)
                        url += "&hasta={}".format(filtro_hasta)
                        filtros = filtros & (Q(fecha_creacion__range=(filtro_desde, filtro_hasta)))
                        # filtros = filtros & (Q(fecha_creacion__gte=filtro_desde, fecha_creacion__lte=filtro_hasta))
                    elif filtro_desde:
                        data['desde_filtrado'] = filtro_desde
                        url += "&desde={}".format(filtro_desde)
                        filtros = filtros & Q(fecha_creacion__gte=filtro_desde)
                    elif filtro_hasta:
                        data['hasta_filtrado'] = filtro_hasta
                        url += "&hasta={}".format(filtro_hasta)
                        filtros = filtros & Q(fecha_creacion__lte=filtro_hasta)
                    else:
                        filtros = filtros & Q(fecha_creacion__gte=datetime.now().date())

                    inscripcioncohorte = InscripcionCohorte.objects.select_related('inscripcionaspirante','inscripcionaspirante__persona','cohortes',
                                                                                   'cohortes__maestriaadmision','cohortes__maestriaadmision__carrera')\
                        .annotate(carrera_nombre= F("cohortes__maestriaadmision__carrera__nombre"),
                                  cohorte_des = F("cohortes__descripcion"),
                                  cedula_per = F("inscripcionaspirante__persona__cedula"),
                                  email_per = F("inscripcionaspirante__persona__email"),
                                  aspirante_per = Concat('inscripcionaspirante__persona__apellido1', V(' '), 'inscripcionaspirante__persona__apellido2', V(' '), 'inscripcionaspirante__persona__nombres'),
                                  telefono_per =F("inscripcionaspirante__persona__telefono") )\
                    .values('carrera_nombre','cohorte_des','fecha_creacion','cedula_per','aspirante_per','email_per','telefono_per')\
                    .filter(filtros)
                    __author__ = 'Unemi'

                    output = io.BytesIO()
                    workbook = xlsxwriter.Workbook(output)
                    ws = workbook.add_worksheet('maestrias')

                    formatoceldagris = workbook.add_format(
                        {'align': 'center', 'border': 1, 'text_wrap': True, 'fg_color': '#B6BFC0'})
                    formatoceldaleft = workbook.add_format({'text_wrap': True, 'align': 'left'})

                    ws.write(0, 0, 'COHORTE', formatoceldagris)
                    ws.write(0, 1, 'MAESTRÍA', formatoceldagris)
                    ws.write(0, 2, 'FEC. CREACIÓN', formatoceldagris)
                    ws.write(0, 3, 'ASPIRANTE', formatoceldagris)
                    ws.write(0, 4, 'CEDULA', formatoceldagris)
                    ws.write(0, 5, 'TELEFONO', formatoceldagris)
                    ws.write(0, 6, 'EMAIL', formatoceldagris)
                    ws.set_column(0, 1, 72)
                    ws.set_column(2, 2, 15)
                    ws.set_column(3, 3, 72)
                    ws.set_column(4, 5, 15)
                    ws.set_column(6, 6, 40)
                    row = 1
                    for ins in inscripcioncohorte:
                        ws.write(row,0,ins['cohorte_des'])
                        ws.write(row,1,ins['carrera_nombre'])
                        ws.write(row,2, str(ins['fecha_creacion'].date()))
                        ws.write(row,3,ins['aspirante_per'])
                        ws.write(row,4,ins['cedula_per'])
                        ws.write(row,5,ins['telefono_per'])
                        ws.write(row,6,ins['email_per'])
                        row +=1
                    workbook.close()
                    output.seek(0)
                    filename = 'maestrias_cohorte.xlsx'  # % (contra.contrato.regimenlaboral)

                    response = HttpResponse(output,
                                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    response['Content-Disposition'] = 'attachment; filename=%s' % filename
                    return response
                except Exception as ex:
                    pass

            elif action == 'matriculadospornivel':
                try:
                    filtro_coordinacion, filtro_nivel, url, filtros = request.GET.get('coordinacion', ''), request.GET.get('nivel', ''), '', Q(status=True)
                    filtrocarreras = Q(status=True)
                    data['nivel_filtrado'] = ''
                    nivel_id = ''
                    if filtro_coordinacion:
                        coordinacion_id = int(filtro_coordinacion)
                        facultad= Coordinacion.objects.get(id=coordinacion_id)
                        filtrocarreras = filtrocarreras & Q(id__in=facultad.carrera.filter(status=True).values_list('id'))
                        url += "&coordinacion={}".format(filtro_coordinacion)
                        data['coordinacion_filtrada'] = int(filtro_coordinacion)

                    if filtro_nivel:
                        data['nivel_filtrado'] = nivel_id = int(filtro_nivel)
                        url += "&nivel={}".format(filtro_nivel)

                    data['url'] = url
                    data['title'] = u'Matriculados por Facultad y Nivel'
                    data['coordinaciones'] = coordina = Coordinacion.objects.filter(status=True, id__in=[1, 2, 3, 4, 5]).order_by('id')
                    data['niveles'] = NivelMalla.objects.filter(status=True).order_by('orden')
                    data['carreras'] = carrera = [c for c in Carrera.objects.filter(filtrocarreras).order_by('id') if c.cantidad_matriculados_nivel(periodo,nivel_id)]
                    data['total_matriculados'] = sum([c.cantidad_matriculados_nivel(periodo,nivel_id) for c in carrera])
                    return render(request, "estadisticas/tablamatriculadospornivel.html", data)

                except Exception as ex:
                    pass

            elif action == 'reporte_matriculadosfacultadnivel':
                try:
                    filtro_coordinacion, filtro_nivel = request.GET.get('coordinacion',''), request.GET.get('nivel','')


                    nombreArchivo = 'MATRICULADOS DE ' + periodo.nombre.__str__() + '.xlsx'

                    libro = xlsxwriter.Workbook(nombreArchivo)
                    ws = libro.add_worksheet()
                    formatoTitulo = libro.add_format({
                        'bold': True,
                        'border': 1,
                        'align': 'center',
                        'valign': 'vcenter',
                    })
                    formatosubtitulo = libro.add_format({
                        'bold': True,
                        'valign': 'vcenter',
                    })

                    ws.merge_range('B1:J1', 'UNIVERSIDAD ESTATAL DE MILAGRO', formatoTitulo)

                    ws.write(4, 0, 'N.', formatosubtitulo)
                    ws.write(4, 1, 'PERIODO', formatosubtitulo)
                    ws.write(4, 2, 'NIVEL_CRE', formatosubtitulo)
                    ws.write(4, 3, 'NIVEL_MAT', formatosubtitulo)
                    ws.write(4, 4, 'SECCION', formatosubtitulo)
                    ws.write(4, 5, 'CEDULA', formatosubtitulo)
                    ws.write(4, 6, 'APELLIDOS', formatosubtitulo)
                    ws.write(4, 7, 'NOMBRES', formatosubtitulo)
                    ws.write(4, 8, 'SEXO', formatosubtitulo)
                    ws.write(4, 9, 'FECHANACIMIENTO', formatosubtitulo)
                    ws.write(4, 10, 'EMAIL', formatosubtitulo)
                    ws.write(4, 11, 'EMAILINST', formatosubtitulo)
                    ws.write(4, 12, 'COORDINACION', formatosubtitulo)
                    ws.write(4, 13, 'CARRERA', formatosubtitulo)
                    ws.write(4, 14, 'COD. SENESCYT', formatosubtitulo)
                    ws.write(4, 15, 'TELEFONO', formatosubtitulo)
                    ws.write(4, 16, 'INSCRIPCION', formatosubtitulo)
                    ws.write(4, 17, '# MATRICULAS', formatosubtitulo)
                    ws.write(4, 18, 'LGTBI', formatosubtitulo)
                    ws.write(4, 19, 'ETNIA', formatosubtitulo)
                    ws.write(4, 20, 'NACIONALIDAD', formatosubtitulo)
                    ws.write(4, 21, 'PAIS', formatosubtitulo)
                    ws.write(4, 22, 'PROVINCIA', formatosubtitulo)
                    ws.write(4, 23, 'CANTON', formatosubtitulo)
                    ws.write(4, 24, 'DIRECCION', formatosubtitulo)
                    ws.write(4, 25, 'ESTADO SOCIO ECONOMICO', formatosubtitulo)
                    ws.write(4, 26, 'REALIZADO', formatosubtitulo)
                    ws.write(4, 27, 'FECHA INICIO PRIMER NIVEL', formatosubtitulo)
                    ws.write(4, 28, 'FECHA CONVALIDACION', formatosubtitulo)
                    ws.write(4, 29, 'TIENE DISCAPACIDAD', formatosubtitulo)
                    ws.write(4, 30, 'DISCAPACIDAD', formatosubtitulo)
                    ws.write(4, 31, 'ID MATRICULA', formatosubtitulo)
                    ws.write(4, 32, 'TIPO MATRICULA', formatosubtitulo)
                    ws.write(4, 33, 'TIPO ESTUDIANTE', formatosubtitulo)
                    ws.write(4, 34, 'CONTACTO EMERGENCIA', formatosubtitulo)
                    ws.write(4, 35, 'COLEGIO', formatosubtitulo)
                    ws.write(4, 36, 'MODALIDAD', formatosubtitulo)
                    ws.write(4, 37, 'PPL', formatosubtitulo)
                    ws.write(4, 38, 'ACEPTA MATRÍCULA', formatosubtitulo)

                    ws.set_column(0, 0, 5)
                    ws.set_column(1, 1, 30)
                    ws.set_column(2, 2, 10)
                    ws.set_column(3, 3, 10)
                    ws.set_column(4, 9, 25)
                    ws.set_column(10, 10, 30)
                    ws.set_column(11, 11, 30)
                    ws.set_column(12, 38, 20)

                    periodoid = str(request.session['periodo'].id)
                    cursor = connections['sga_select'].cursor()

                    if filtro_coordinacion and filtro_nivel:
                        filtro_coordinacion = int(filtro_coordinacion)
                        filtro_nivel = int(filtro_nivel)
                        listaestudiante = """select peri.nombre as PERIODO, nimalla.nombre as NIVEL_CRE,
                                                            sesion.nombre as SECCION,
                                                            perso.cedula, perso.apellido1 || ' ' || perso.apellido2 as apellidos, 
                                                            perso.nombres as nompersona, se.nombre as sexo, 
                                                            perso.nacimiento as fechanacimiento,perso.email,perso.emailinst, coor.nombre as facultad, carr.nombre as carrera, 
                                                            (select ma.codigo from sga_inscripcionmalla insmalla inner join sga_malla ma on insmalla.malla_id=ma.id where insmalla.inscripcion_id=ins.id and insmalla.status=True), 
                                                            perso.telefono,ins.id as INSCRIPCION, 
                                                            CASE WHEN perso.lgtbi = True THEN 'SI' else 'NO' END as lgtbi, 
                                                            (select rasa.nombre from sga_perfilinscripcion perfil inner join sga_raza rasa on perfil.raza_id=rasa.id where perfil.persona_id=perso.id and perfil.status=True) as etnia, 
                                                            perso.nacionalidad, pais.nombre as pais, provincia.nombre as provincia, canton.nombre as canton, 
                                                            perso.direccion || ' ' || perso.direccion2 as direccion, 
                                                            (select gru.codigo || ' ' || gru.nombre from socioecon_fichasocioeconomicainec fi inner join socioecon_gruposocioeconomico gru on fi.grupoeconomico_id=gru.id where fi.persona_id=perso.id and fi.status=True) as gruposocioeconomico, 
                                                            ins.fechainicioprimernivel,ins.fechainicioconvalidacion, 
                                                            (select CASE WHEN perfilin.tienediscapacidad = True THEN 'SI' else 'NO' END  from sga_perfilinscripcion perfilin where perfilin.persona_id=perso.id and perfilin.status=True) as tienediscapacidad, 
                                                            (select disca.nombre  from sga_perfilinscripcion perfilin,sga_discapacidad disca where perfilin.persona_id=perso.id and perfilin.tipodiscapacidad_id=disca.id and perfilin.status=True) as discapacidad, 
                                                            matri.id as idmatricula, CASE WHEN gruso.tipomatricula=1 THEN 'REGULAR' WHEN gruso.tipomatricula=2 THEN 'IRREGULAR' END as tipomatricula, 
                                                            tima.nombre as tipoestudiante, ext.contactoemergencia as contactoemergencia, 
                                                            uni.nombre as unidadeducativa, modal.nombre as modalidad,
                                                            (select count(*) from sga_matricula matr where matr.inscripcion_id=ins.id and matr.status=True
                                                            and matr.retiradomatricula=False) as veces_matricula , perso.ppl, matri.termino
                                                            from sga_matricula matri                                   


                                                            inner join sga_tipomatricula tima on tima.id=matri.tipomatricula_id and tima.status=True

                                                            inner join sga_nivel ni on matri.nivel_id=ni.id and ni.status=True


                                                            INNER JOIN sga_nivellibrecoordinacion  ON ni.id = sga_nivellibrecoordinacion.nivel_id                                  

                                                            inner join sga_inscripcion ins on matri.inscripcion_id=ins.id and ins.status=True
                                                            left join sga_coordinacion coor on coor.id=ins.coordinacion_id                                 

                                                            left join sga_carrera carr on carr.id=ins.carrera_id
                                                            left join sga_institucionescolegio uni on uni.id=ins.unidadeducativa_id and uni.status=True
                                                            left join sga_matriculagruposocioeconomico gruso on gruso.matricula_id=matri.id and gruso.status=True
                                                            inner join sga_persona perso on ins.persona_id=perso.id
                                                            left join med_personaextension ext on ext.persona_id=perso.id and ext.status=True
                                                            left join sga_sexo se on se.id=perso.sexo_id and se.status=True
                                                            left join sga_pais pais on perso.pais_id=pais.id and pais.status=True
                                                            left join sga_provincia provincia on perso.provincia_id=provincia.id and provincia.status=True
                                                            left join sga_canton canton on perso.canton_id=canton.id and canton.status=True
                                                            inner join sga_periodo peri on ni.periodo_id=peri.id
                                                            inner join sga_nivelmalla nimalla on matri.nivelmalla_id=nimalla.id 
                                                            left join sga_sesion sesion on ins.sesion_id=sesion.id
                                                            inner join sga_modalidad modal on ins.modalidad_id=modal.id
                                                            where  matri.status=True and matri.retiradomatricula=False
                                                            and ni.periodo_id= 153 
                                                            AND coor.id = %s
                        									AND matri.nivelmalla_id=%s	
                        									and matri.estado_matricula IN (2, 3) ORDER BY apellidos""" % (filtro_coordinacion, filtro_nivel)

                    elif filtro_coordinacion:
                        filtro_coordinacion = int(filtro_coordinacion)
                        listaestudiante = """select peri.nombre as PERIODO, nimalla.nombre as NIVEL_CRE,
                                                                                    sesion.nombre as SECCION,
                                                                                    perso.cedula, perso.apellido1 || ' ' || perso.apellido2 as apellidos, 
                                                                                    perso.nombres as nompersona, se.nombre as sexo, 
                                                                                    perso.nacimiento as fechanacimiento,perso.email,perso.emailinst, coor.nombre as facultad, carr.nombre as carrera, 
                                                                                    (select ma.codigo from sga_inscripcionmalla insmalla inner join sga_malla ma on insmalla.malla_id=ma.id where insmalla.inscripcion_id=ins.id and insmalla.status=True), 
                                                                                    perso.telefono,ins.id as INSCRIPCION, 
                                                                                    CASE WHEN perso.lgtbi = True THEN 'SI' else 'NO' END as lgtbi, 
                                                                                    (select rasa.nombre from sga_perfilinscripcion perfil inner join sga_raza rasa on perfil.raza_id=rasa.id where perfil.persona_id=perso.id and perfil.status=True) as etnia, 
                                                                                    perso.nacionalidad, pais.nombre as pais, provincia.nombre as provincia, canton.nombre as canton, 
                                                                                    perso.direccion || ' ' || perso.direccion2 as direccion, 
                                                                                    (select gru.codigo || ' ' || gru.nombre from socioecon_fichasocioeconomicainec fi inner join socioecon_gruposocioeconomico gru on fi.grupoeconomico_id=gru.id where fi.persona_id=perso.id and fi.status=True) as gruposocioeconomico, 
                                                                                    ins.fechainicioprimernivel,ins.fechainicioconvalidacion, 
                                                                                    (select CASE WHEN perfilin.tienediscapacidad = True THEN 'SI' else 'NO' END  from sga_perfilinscripcion perfilin where perfilin.persona_id=perso.id and perfilin.status=True) as tienediscapacidad, 
                                                                                    (select disca.nombre  from sga_perfilinscripcion perfilin,sga_discapacidad disca where perfilin.persona_id=perso.id and perfilin.tipodiscapacidad_id=disca.id and perfilin.status=True) as discapacidad, 
                                                                                    matri.id as idmatricula, CASE WHEN gruso.tipomatricula=1 THEN 'REGULAR' WHEN gruso.tipomatricula=2 THEN 'IRREGULAR' END as tipomatricula, 
                                                                                    tima.nombre as tipoestudiante, ext.contactoemergencia as contactoemergencia, 
                                                                                    uni.nombre as unidadeducativa, modal.nombre as modalidad,
                                                                                    (select count(*) from sga_matricula matr where matr.inscripcion_id=ins.id and matr.status=True
                                                                                    and matr.retiradomatricula=False) as veces_matricula , perso.ppl, matri.termino
                                                                                    from sga_matricula matri                                   


                                                                                    inner join sga_tipomatricula tima on tima.id=matri.tipomatricula_id and tima.status=True

                                                                                    inner join sga_nivel ni on matri.nivel_id=ni.id and ni.status=True


                                                                                    INNER JOIN sga_nivellibrecoordinacion  ON ni.id = sga_nivellibrecoordinacion.nivel_id                                  

                                                                                    inner join sga_inscripcion ins on matri.inscripcion_id=ins.id and ins.status=True
                                                                                    left join sga_coordinacion coor on coor.id=ins.coordinacion_id                                 

                                                                                    left join sga_carrera carr on carr.id=ins.carrera_id
                                                                                    left join sga_institucionescolegio uni on uni.id=ins.unidadeducativa_id and uni.status=True
                                                                                    left join sga_matriculagruposocioeconomico gruso on gruso.matricula_id=matri.id and gruso.status=True
                                                                                    inner join sga_persona perso on ins.persona_id=perso.id
                                                                                    left join med_personaextension ext on ext.persona_id=perso.id and ext.status=True
                                                                                    left join sga_sexo se on se.id=perso.sexo_id and se.status=True
                                                                                    left join sga_pais pais on perso.pais_id=pais.id and pais.status=True
                                                                                    left join sga_provincia provincia on perso.provincia_id=provincia.id and provincia.status=True
                                                                                    left join sga_canton canton on perso.canton_id=canton.id and canton.status=True
                                                                                    inner join sga_periodo peri on ni.periodo_id=peri.id
                                                                                    inner join sga_nivelmalla nimalla on matri.nivelmalla_id=nimalla.id 
                                                                                    left join sga_sesion sesion on ins.sesion_id=sesion.id
                                                                                    inner join sga_modalidad modal on ins.modalidad_id=modal.id
                                                                                    where  matri.status=True and matri.retiradomatricula=False
                                                                                    and ni.periodo_id= 153 
                                                                                    AND coor.id = %s
                                                									and matri.estado_matricula IN (2, 3) ORDER BY apellidos""" % (filtro_coordinacion)

                    elif filtro_nivel:
                        filtro_nivel = int(filtro_nivel)
                        listaestudiante = """select peri.nombre as PERIODO, nimalla.nombre as NIVEL_CRE,
                                                            sesion.nombre as SECCION,
                                                            perso.cedula, perso.apellido1 || ' ' || perso.apellido2 as apellidos, 
                                                            perso.nombres as nompersona, se.nombre as sexo, 
                                                            perso.nacimiento as fechanacimiento,perso.email,perso.emailinst, coor.nombre as facultad, carr.nombre as carrera, 
                                                            (select ma.codigo from sga_inscripcionmalla insmalla inner join sga_malla ma on insmalla.malla_id=ma.id where insmalla.inscripcion_id=ins.id and insmalla.status=True), 
                                                            perso.telefono,ins.id as INSCRIPCION, 
                                                            CASE WHEN perso.lgtbi = True THEN 'SI' else 'NO' END as lgtbi, 
                                                            (select rasa.nombre from sga_perfilinscripcion perfil inner join sga_raza rasa on perfil.raza_id=rasa.id where perfil.persona_id=perso.id and perfil.status=True) as etnia, 
                                                            perso.nacionalidad, pais.nombre as pais, provincia.nombre as provincia, canton.nombre as canton, 
                                                            perso.direccion || ' ' || perso.direccion2 as direccion, 
                                                            (select gru.codigo || ' ' || gru.nombre from socioecon_fichasocioeconomicainec fi inner join socioecon_gruposocioeconomico gru on fi.grupoeconomico_id=gru.id where fi.persona_id=perso.id and fi.status=True) as gruposocioeconomico, 
                                                            ins.fechainicioprimernivel,ins.fechainicioconvalidacion, 
                                                            (select CASE WHEN perfilin.tienediscapacidad = True THEN 'SI' else 'NO' END  from sga_perfilinscripcion perfilin where perfilin.persona_id=perso.id and perfilin.status=True) as tienediscapacidad, 
                                                            (select disca.nombre  from sga_perfilinscripcion perfilin,sga_discapacidad disca where perfilin.persona_id=perso.id and perfilin.tipodiscapacidad_id=disca.id and perfilin.status=True) as discapacidad, 
                                                            matri.id as idmatricula, CASE WHEN gruso.tipomatricula=1 THEN 'REGULAR' WHEN gruso.tipomatricula=2 THEN 'IRREGULAR' END as tipomatricula, 
                                                            tima.nombre as tipoestudiante, ext.contactoemergencia as contactoemergencia, 
                                                            uni.nombre as unidadeducativa, modal.nombre as modalidad,
                                                            (select count(*) from sga_matricula matr where matr.inscripcion_id=ins.id and matr.status=True
                                                            and matr.retiradomatricula=False) as veces_matricula , perso.ppl, matri.termino
                                                            from sga_matricula matri                                   


                                                            inner join sga_tipomatricula tima on tima.id=matri.tipomatricula_id and tima.status=True

                                                            inner join sga_nivel ni on matri.nivel_id=ni.id and ni.status=True


                                                            INNER JOIN sga_nivellibrecoordinacion  ON ni.id = sga_nivellibrecoordinacion.nivel_id                                  

                                                            inner join sga_inscripcion ins on matri.inscripcion_id=ins.id and ins.status=True
                                                            left join sga_coordinacion coor on coor.id=ins.coordinacion_id                                 

                                                            left join sga_carrera carr on carr.id=ins.carrera_id
                                                            left join sga_institucionescolegio uni on uni.id=ins.unidadeducativa_id and uni.status=True
                                                            left join sga_matriculagruposocioeconomico gruso on gruso.matricula_id=matri.id and gruso.status=True
                                                            inner join sga_persona perso on ins.persona_id=perso.id
                                                            left join med_personaextension ext on ext.persona_id=perso.id and ext.status=True
                                                            left join sga_sexo se on se.id=perso.sexo_id and se.status=True
                                                            left join sga_pais pais on perso.pais_id=pais.id and pais.status=True
                                                            left join sga_provincia provincia on perso.provincia_id=provincia.id and provincia.status=True
                                                            left join sga_canton canton on perso.canton_id=canton.id and canton.status=True
                                                            inner join sga_periodo peri on ni.periodo_id=peri.id
                                                            inner join sga_nivelmalla nimalla on matri.nivelmalla_id=nimalla.id 
                                                            left join sga_sesion sesion on ins.sesion_id=sesion.id
                                                            inner join sga_modalidad modal on ins.modalidad_id=modal.id
                                                            where  matri.status=True and matri.retiradomatricula=False
                                                            and ni.periodo_id= 153 AND matri.nivelmalla_id=%s	
                        									and matri.estado_matricula IN (2, 3) ORDER BY apellidos""" % (filtro_nivel)

                    else:
                        listaestudiante = """select peri.nombre as PERIODO, nimalla.nombre as NIVEL_CRE,
                                                            sesion.nombre as SECCION,
                                                            perso.cedula, perso.apellido1 || ' ' || perso.apellido2 as apellidos, 
                                                            perso.nombres as nompersona, se.nombre as sexo, 
                                                            perso.nacimiento as fechanacimiento,perso.email,perso.emailinst, coor.nombre as facultad, carr.nombre as carrera, 
                                                            (select ma.codigo from sga_inscripcionmalla insmalla inner join sga_malla ma on insmalla.malla_id=ma.id where insmalla.inscripcion_id=ins.id and insmalla.status=True), 
                                                            perso.telefono,ins.id as INSCRIPCION, 
                                                            CASE WHEN perso.lgtbi = True THEN 'SI' else 'NO' END as lgtbi, 
                                                            (select rasa.nombre from sga_perfilinscripcion perfil inner join sga_raza rasa on perfil.raza_id=rasa.id where perfil.persona_id=perso.id and perfil.status=True) as etnia, 
                                                            perso.nacionalidad, pais.nombre as pais, provincia.nombre as provincia, canton.nombre as canton, 
                                                            perso.direccion || ' ' || perso.direccion2 as direccion, 
                                                            (select gru.codigo || ' ' || gru.nombre from socioecon_fichasocioeconomicainec fi inner join socioecon_gruposocioeconomico gru on fi.grupoeconomico_id=gru.id where fi.persona_id=perso.id and fi.status=True) as gruposocioeconomico, 
                                                            ins.fechainicioprimernivel,ins.fechainicioconvalidacion, 
                                                            (select CASE WHEN perfilin.tienediscapacidad = True THEN 'SI' else 'NO' END  from sga_perfilinscripcion perfilin where perfilin.persona_id=perso.id and perfilin.status=True) as tienediscapacidad, 
                                                            (select disca.nombre  from sga_perfilinscripcion perfilin,sga_discapacidad disca where perfilin.persona_id=perso.id and perfilin.tipodiscapacidad_id=disca.id and perfilin.status=True) as discapacidad, 
                                                            matri.id as idmatricula, CASE WHEN gruso.tipomatricula=1 THEN 'REGULAR' WHEN gruso.tipomatricula=2 THEN 'IRREGULAR' END as tipomatricula, 
                                                            tima.nombre as tipoestudiante, ext.contactoemergencia as contactoemergencia, 
                                                            uni.nombre as unidadeducativa, modal.nombre as modalidad,
                                                            (select count(*) from sga_matricula matr where matr.inscripcion_id=ins.id and matr.status=True
                                                            and matr.retiradomatricula=False) as veces_matricula , perso.ppl, matri.termino
                                                            from sga_matricula matri                                   


                                                            inner join sga_tipomatricula tima on tima.id=matri.tipomatricula_id and tima.status=True

                                                            inner join sga_nivel ni on matri.nivel_id=ni.id and ni.status=True


                                                            INNER JOIN sga_nivellibrecoordinacion  ON ni.id = sga_nivellibrecoordinacion.nivel_id                                  

                                                            inner join sga_inscripcion ins on matri.inscripcion_id=ins.id and ins.status=True
                                                            left join sga_coordinacion coor on coor.id=ins.coordinacion_id                                 

                                                            left join sga_carrera carr on carr.id=ins.carrera_id
                                                            left join sga_institucionescolegio uni on uni.id=ins.unidadeducativa_id and uni.status=True
                                                            left join sga_matriculagruposocioeconomico gruso on gruso.matricula_id=matri.id and gruso.status=True
                                                            inner join sga_persona perso on ins.persona_id=perso.id
                                                            left join med_personaextension ext on ext.persona_id=perso.id and ext.status=True
                                                            left join sga_sexo se on se.id=perso.sexo_id and se.status=True
                                                            left join sga_pais pais on perso.pais_id=pais.id and pais.status=True
                                                            left join sga_provincia provincia on perso.provincia_id=provincia.id and provincia.status=True
                                                            left join sga_canton canton on perso.canton_id=canton.id and canton.status=True
                                                            inner join sga_periodo peri on ni.periodo_id=peri.id
                                                            inner join sga_nivelmalla nimalla on matri.nivelmalla_id=nimalla.id 
                                                            left join sga_sesion sesion on ins.sesion_id=sesion.id
                                                            inner join sga_modalidad modal on ins.modalidad_id=modal.id
                                                            where  matri.status=True and matri.retiradomatricula=False
                                                            and ni.periodo_id= 153 and matri.estado_matricula IN (2, 3) ORDER BY apellidos""" % (periodoid)


                    cursor.execute(listaestudiante)
                    results = cursor.fetchall()
                    a = 4

                    for per in results:
                        a += 1
                        ws.write(a, 0, a - 4)
                        ws.write(a, 1, '%s' % per[0])
                        ws.write(a, 2, '%s' % per[1])
                        ws.write(a, 3, '%s' % per[1])
                        ws.write(a, 4, '%s' % per[2])
                        ws.write(a, 5, '%s' % per[3])
                        ws.write(a, 6, '%s' % per[4])
                        ws.write(a, 7, '%s' % per[5])
                        ws.write(a, 8, '%s' % per[6])
                        ws.write(a, 9, '%s' % per[7])
                        # ws.write(a, 9, per[7], date_format)
                        ws.write(a, 10, '%s' % per[8])
                        ws.write(a, 11, '%s' % per[9])
                        ws.write(a, 12, '%s' % per[10])
                        ws.write(a, 13, '%s' % per[11])
                        ws.write(a, 14, '%s' % per[12])
                        ws.write(a, 15, '%s' % per[13])
                        ws.write(a, 16, '%s' % per[14])
                        ws.write(a, 17, '%s' % per[33])
                        ws.write(a, 18, '%s' % per[15])
                        ws.write(a, 19, '%s' % per[16])
                        ws.write(a, 20, '%s' % per[17])
                        ws.write(a, 21, '%s' % per[18])
                        ws.write(a, 22, '%s' % per[19])
                        ws.write(a, 23, '%s' % per[20])
                        ws.write(a, 24, '%s' % per[21])
                        ws.write(a, 25, '%s' % per[22])
                        ws.write(a, 26, 'SI')
                        ws.write(a, 27, '%s' % per[23])
                        ws.write(a, 28, '%s' % per[24])
                        ws.write(a, 29, '%s' % per[25])
                        ws.write(a, 30, '%s' % per[26])
                        ws.write(a, 31, '%s' % per[27])
                        ws.write(a, 32, '%s' % per[28])
                        ws.write(a, 33, '%s' % per[29])
                        ws.write(a, 34, '%s' % per[30])
                        ws.write(a, 35, '%s' % per[31])
                        ws.write(a, 36, '%s' % per[32])
                        ws.write(a, 37, '%s' % "SI" if per[34] else "NO")
                        ws.write(a, 38, '%s' % "SI" if per[35] else "NO")

                    libro.close()
                    # Crear respuesta HttpResponse
                    response = HttpResponse(content_type="application/ms-excel")
                    content = "attachment; filename = {0}".format(nombreArchivo)
                    response['Content-Disposition'] = content
                    response.write(open(nombreArchivo, 'rb').read())

                    return response

                except Exception as ex:
                    pass


            elif action == 'aspirantesmaestrias':
                try:
                    data['title'] = u'Aspirantes de Maestrías'
                    print(request.GET)
                    filtro_carrera, filtro_desde, filtros, filtros_c, url = request.GET.get('carrera', ''), request.GET.get('desde', ''),  Q(status=True), Q(status=True), ''
                    filtro_hasta=None
                    if request.GET.get('hasta', ''):
                        filtro_hasta=u"%s %s"% (request.GET.get('hasta', ''),'23:59:59')
                    if filtro_carrera:
                        data['carrera_filtrada'] = carrera_id = int(filtro_carrera)
                        url += "&carrera={}".format(filtro_carrera)
                        filtros = filtros & Q(cohortes__maestriaadmision__carrera_id=carrera_id)
                    # else:
                    #     filtros_c = filtros_c & Q(id)
                    if filtro_desde and filtro_hasta:
                        data['desde_filtrado'] = filtro_desde
                        data['hasta_filtrado'] = request.GET.get('hasta', '')
                        url += "&desde={}".format(filtro_desde)
                        url += "&hasta={}".format(filtro_hasta)
                        filtros = filtros & (Q(fecha_creacion__range=(filtro_desde, filtro_hasta)))
                        # filtros = filtros & (Q(fecha_creacion__gte=filtro_desde, fecha_creacion__lte=filtro_hasta))
                    elif filtro_desde:
                        data['desde_filtrado'] = filtro_desde
                        url += "&desde={}".format(filtro_desde)
                        filtros = filtros & Q(fecha_creacion__gte=filtro_desde)
                    elif filtro_hasta:
                        data['hasta_filtrado'] = filtro_hasta
                        url += "&hasta={}".format(filtro_hasta)
                        filtros = filtros & Q(fecha_creacion__lte=filtro_hasta)
                    else:
                        filtros = filtros & Q(fecha_creacion__gte=datetime.now().date())

                    carrerasmaestrias = Carrera.objects.filter(status=True, id__in=MaestriasAdmision.objects.filter(status=True).values_list('carrera_id', flat=True))
                    data['inscritos_cohorte'] = inscripcioncohorte = InscripcionCohorte.objects.filter(filtros)
                    data['cantidad_inscritos_cohorte'] = len(inscripcioncohorte)
                    cantidad_aspirantes_anio = []
                    cantidad_aspirantes_mes = []
                    cantidad_aspirantes_dia = []
                    cantidad_aspirantes_carrera = []
                    anios_maestrias = []
                    meses_maestrias = []
                    dias_maestrias = []
                    for anio in inscripcioncohorte.annotate(anio=ExtractYear('fecha_creacion')).values_list('anio', flat=True).order_by('anio').distinct():
                        aspirantes = len(inscripcioncohorte.filter(fecha_creacion__year = anio))
                        cantidad_aspirantes_anio.append(aspirantes)
                        anios_maestrias.append(anio)
                    meseslist = inscripcioncohorte.annotate(mes=ExtractMonth('fecha_creacion')).filter(fecha_creacion__year__in=anios_maestrias).values_list('mes', flat=True).order_by('mes').distinct()
                    for mes in meseslist:
                        aspirantes = len(inscripcioncohorte.filter(fecha_creacion__month=mes))
                        cantidad_aspirantes_mes.append(aspirantes)
                        meses_maestrias.append(str(transformar_mes(mes)))

                    # for dia in inscripcioncohorte.filter(fecha_creacion__month__in=meseslist).values_list()

                    # inscripcioncohorte.annotate(dia=ExtractDay('fecha_creacion')).filter(fecha_creacion__month__in=meseslist).values_list('dia', flat=True).order_by('-dia').distinct()
                    for dia in inscripcioncohorte.annotate(dia=ExtractDay('fecha_creacion'), mes=ExtractMonth('fecha_creacion'), anio=ExtractYear('fecha_creacion')).filter(fecha_creacion__month__in=meseslist).values_list('dia', 'mes', 'anio').order_by('mes', 'anio').distinct():
                        aspirantes = len(inscripcioncohorte.filter(fecha_creacion__day=dia[0]))
                        cantidad_aspirantes_dia.append(aspirantes)
                        # dia_le=diaenletra(dia)
                        fecha = datetime(dia[2], dia[1], dia[0])
                        fecha_registro = '{} {} {}'.format(transformar_mes(fecha.month), fecha.day, fecha.year)
                        dias_maestrias.append(fecha_registro)

                    for carrera in carrerasmaestrias:
                        aspirantes = len(inscripcioncohorte.filter(filtros, cohortes__maestriaadmision__carrera=carrera))
                        cantidad_aspirantes_carrera.append(aspirantes)

                    data['cantidad_aspirantes_anio'] = cantidad_aspirantes_anio
                    data['cantidad_aspirantes_mes'] = cantidad_aspirantes_mes
                    data['cantidad_aspirantes_dias'] = cantidad_aspirantes_dia
                    data['cantidad_aspirantes_carrera'] = cantidad_aspirantes_carrera
                    data['anios_maestrias'] = anios_maestrias
                    data['meses_maestrias'] = meses_maestrias
                    data['dias_maestrias'] = dias_maestrias
                    data['carreras_maestrias'] = list(carrerasmaestrias.values_list('nombre', flat=True))
                    data['carreras'] = carrerasmaestrias
                    data['desde'] = filtro_desde
                    data['hasta'] = filtro_hasta
                    data['url'] = url
                    return render(request, "estadisticas/aspirantesmaestrias.html", data)
                except Exception as ex:
                    print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
                    pass

            elif action == 'prematriculacarrera':
                try:
                    data['title'] = u'Pre-Matriculados y Matriculados por carreras'

                    querypre = PreMatricula.objects.filter(status=True, periodo=periodo)
                    querymat = Matricula.objects.filter(estado_matricula__in=[2, 3], nivel__periodo=periodo, status=True)
                    coordpremat = querypre.values_list('inscripcion__carrera__coordinacion', flat=True).distinct()
                    coordmat = querymat.values_list('inscripcion__carrera__coordinacion', flat=True).distinct()
                    querysesion = Sesion.objects.filter(Q(status=True), Q(id__in=[1, 4, 5, 7, 13]))
                    data['coordinaciones'] = coordinaciones = Coordinacion.objects.filter(Q(status=True), Q(id__in=coordmat) | Q(id__in=coordpremat)).distinct().values_list('id', 'nombre')
                    coordinacion = 0
                    carrera = 0
                    if 'coordinacion' in request.GET and not request.GET['coordinacion']=='':
                        coordinacion = int(request.GET['coordinacion'])

                    data['coordinacion'] = coordinacion
                    coordinacionfilter = coordinaciones.values_list('id',flat=True) if coordinacion == 0 else [coordinacion]


                    query = Carrera.objects.filter(status=True, coordinacion__in=coordinacionfilter).distinct().order_by('coordinacion', 'nombre')

                    data['carrerasfilter'] = [c for c in query if c.cantidad_matriculados(periodo) or c.cantidad_prematriculados_sesion(periodo, [1, 4, 5, 7, 13])]

                    if 'carrera' in request.GET and not request.GET['carrera'] == '' and not request.GET['carrera'] == '0':
                        data['carrera'] = carrera = int(request.GET['carrera'])
                        query = query.filter(id=carrera)

                    if 'sesion' in request.GET and not request.GET['sesion'] == '' and not request.GET['sesion'] == '0':
                        data['sesion'] = sesion = int(request.GET['sesion'])
                        data['sesiones'] = sesiones = querysesion.filter(Q(id=sesion)).order_by('id')
                    else:
                        data['sesiones'] = sesiones = querysesion.order_by('id')
                    data['colspan'] = len(sesiones) + 1
                    data['colspang'] = ((len(sesiones) + 1) * 2) + 1
                    data['sesiones_ids'] = sesiones.values_list('id', flat=True)

                    if carrera > 0:
                        data['sesionfilter'] = querysesion.filter(Q(id__in=querypre.filter(inscripcion__carrera__coordinacion__in=coordinacionfilter, inscripcion__carrera__id=carrera).values_list('sesion_id', flat=True).distinct()) |
                                                                Q(id__in=querymat.filter(inscripcion__carrera__coordinacion__in=coordinacionfilter, inscripcion__carrera__id=carrera).values_list('nivel__sesion_id', flat=True).distinct())).order_by('id')
                    else:
                        data['sesionfilter'] = querysesion.filter(Q(id__in=querypre.filter(inscripcion__carrera__coordinacion__in=coordinacionfilter).values_list('sesion_id', flat=True).distinct()) |
                                                              Q(id__in=querymat.filter(inscripcion__carrera__coordinacion__in=coordinacionfilter).values_list('nivel__sesion_id', flat=True).distinct())).order_by('id')

                    data['carreras'] = [c for c in query if c.cantidad_matriculados(periodo) or c.cantidad_prematriculados_sesion(periodo, [1, 4, 5, 7, 13])]

                    # data['total_matriculados_mujeres'] = total_matriculados_mujeres(periodo)
                    # data['total_matriculados_hombres'] = total_matriculados_hombres(periodo)
                    # data['total_matriculados'] = total_matriculados(periodo)
                    # data['total_prematriculados'] = total_prematriculados(periodo)
                    # data['total_prematriculados_sesion'] = total_prematriculados_sesion(periodo)
                    # data['total_prematriculados'] = total_prematriculados(periodo)
                    return render(request, "estadisticas/estadisticasprematricula.html", data)
                except Exception as ex:
                    pass

            elif action == 'carrerascorrdinacion':
                try:
                    data = {}
                    querypre = PreMatricula.objects.filter(status=True, periodo=periodo)
                    querymat = Matricula.objects.filter(estado_matricula__in=[2, 3], nivel__periodo=periodo, status=True)
                    coordpremat = querypre.values_list('inscripcion__carrera__coordinacion', flat=True).distinct()
                    coordmat = querymat.values_list('inscripcion__carrera__coordinacion', flat=True).distinct()
                    querysesion = Sesion.objects.filter(Q(status=True), Q(id__in=[1, 4, 5, 7, 13]))
                    coordinaciones = Coordinacion.objects.filter(Q(status=True), Q(id__in=coordmat) | Q(id__in=coordpremat)).distinct().values_list('id', 'nombre')
                    coordinacion = 0
                    if 'coordinacion' in request.GET and not request.GET['coordinacion'] == '':
                        coordinacion = int(request.GET['coordinacion'])

                    data['coordinacion'] = coordinacion
                    coordinacionfilter = coordinaciones.values_list('id', flat=True) if coordinacion == 0 else [coordinacion]

                    data['lista'] = [{'nombre': c.nombre, 'id': c.id } for c in Carrera.objects.filter(status=True, coordinacion__in=coordinacionfilter).distinct().order_by('coordinacion', 'nombre') if c.cantidad_matriculados(periodo) or c.cantidad_prematriculados_sesion(periodo, [1, 3, 4, 7, 13])]

                    sesiones = querysesion.filter(Q(id__in=querypre.filter(inscripcion__carrera__coordinacion__in=coordinacionfilter).values_list('sesion_id', flat=True).distinct()) |
                                                              Q(id__in=querymat.filter(inscripcion__carrera__coordinacion__in=coordinacionfilter).values_list('nivel__sesion_id', flat=True).distinct())).order_by('id')
                    data['sesiones'] = [{'nombre': s.nombre, 'id': s.id} for s in sesiones]
                    data['result'] = 'ok'
                    return JsonResponse(data)
                except Exception as e:
                    print(e)
                    transaction.set_rollback(True)
                    pass

            elif action == 'estadisticacarrera':
                try:
                    data = {}
                    data['periodo'] = periodo
                    data['carrera'] = carrera = Carrera.objects.get(id=request.GET['carrera'])

                    querypre = PreMatricula.objects.filter(status=True, periodo=periodo)
                    querymat = Matricula.objects.filter(estado_matricula__in=[2, 3], nivel__periodo=periodo, status=True)
                    data['sesiones'] = sesiones = Sesion.objects.filter(Q(id__in=querypre.filter(inscripcion__carrera=carrera).values_list('sesion_id', flat=True).distinct()) |Q(id__in=querymat.filter(inscripcion__carrera=carrera).values_list('nivel__sesion_id', flat=True).distinct())).order_by('id')
                    data['sesiones_ids'] = sesiones.values_list('id', flat=True)
                    data['colspan'] = len(sesiones) + 1
                    data['colspang'] = ((len(sesiones) + 1) * 2) + 1
                    titulo = 'ESTADÍSTICAS DE LA CARRERA {}'.format(carrera)
                    template = get_template("estadisticas/carrerasestadistica.html")
                    return JsonResponse({"result": 'ok', 'cuerpo': template.render(data), 'titulo': titulo})
                except Exception as e:
                    print(e)
                    pass

            elif action == 'segmentobienestarmatriculados':
                try:
                    data['title'] = u'Matriculados por coordinaciones segun su tipos de matricula'
                    data['coordinaciones'] = [c for c in Coordinacion.objects.all().order_by('id') if c.cantidad_matriculados(periodo)]
                    matriculados = Matricula.objects.filter(status=True, nivel__periodo_id=periodo.id)
                    matriculados_cal = matriculados.aggregate(total_matriculados_regulares=Count('matriculagruposocioeconomico__id', filter=Q(matriculagruposocioeconomico__tipomatricula=1, matriculagruposocioeconomico__status=True)),
                                                              total_matriculados_irregulares=Count('matriculagruposocioeconomico__id', filter=Q(matriculagruposocioeconomico__tipomatricula=2, matriculagruposocioeconomico__status=True)),
                                                              total_matriculados_regulares_retirados=Count('matriculagruposocioeconomico__id', filter=Q(matriculagruposocioeconomico__tipomatricula=1, retiradomatricula=True, matriculagruposocioeconomico__status=True)),
                                                              total_matriculados_regulares_no_retirados=Count('matriculagruposocioeconomico__id', filter=Q(matriculagruposocioeconomico__tipomatricula=1, retiradomatricula=False, matriculagruposocioeconomico__status=True)),
                                                              total_matriculados_irregulares_retirados=Count('matriculagruposocioeconomico__id', filter=Q(matriculagruposocioeconomico__tipomatricula=2, retiradomatricula=True, matriculagruposocioeconomico__status=True)),
                                                              total_matriculados_irregulares_no_retirados=Count('matriculagruposocioeconomico__id', filter=Q(matriculagruposocioeconomico__tipomatricula=2, retiradomatricula=False, matriculagruposocioeconomico__status=True)),
                                                              total_matriculados=Count('matriculagruposocioeconomico__id', filter=Q(matriculagruposocioeconomico__status=True)),)
                    data.update(matriculados_cal)
                    return render(request, "estadisticas/tablasegmentobienestarmatriculados.html", data)
                except Exception as ex:
                    pass

            elif action == 'segmentobienestarpreseleccionados':
                try:
                    data['title'] = u'Matriculados por coordinaciones segun su tipos de matricula'
                    preinscripciones = PreInscripcionBeca.objects.filter(status=True, periodo_id=periodo.id)
                    becastipos_ids = preinscripciones.values_list('becatipo_id', flat=True).distinct()
                    data['becastipos'] = BecaTipo.objects.filter(id__in=becastipos_ids, status=True)
                    data['totales'] = preinscripciones.count()

                    return render(request, "estadisticas/tablasegmentobienestarpreseleccionados.html", data)
                except Exception as ex:
                    pass

            elif action == 'segmentobienestarajudicados':
                try:
                    data['title'] = u'Ajudicación de becas  segun su tipos de beca'
                    data['becassolicitudes'] = becassolicitudes = BecaSolicitud.objects.filter(status=True, periodo_id=periodo.id)
                    becastipos_ids = becassolicitudes.values_list('becatipo_id', flat=True).distinct()


                    becassolicitudes_ids = becassolicitudes.values_list('id', flat=True)
                    data['becasasignaciones'] = becasasignaciones = BecaAsignacion.objects.filter(status=True,
                                                                                                  solicitud_id__in=becassolicitudes_ids)
                    estados = dict(ESTADO_SOLICITUD_BECAS)
                    estados[0] = 'TOTAL'
                    data['becasestados'] = estados.items()
                    subquery = BecaSolicitudRecorrido.objects.filter(solicitud=OuterRef('pk'), status=True).order_by('-fecha_creacion')
                    data['totalesbecasestados'] = [becassolicitudes.annotate(estado_ultimo=Subquery(subquery.values('estado')[:1])).filter(estado_ultimo__in=estados.keys()if key == 0 else [key]).count()for key, value in data['becasestados']]
                   # data['totalesbecasestados'] = [becassolicitudes.filter(estado__in=estados.keys() if key == 0 else [key], status=True).count()for key, value in data['becasestados']]

                    estadosaceptados= dict(ESTADO_ACEPTACION_BECA)
                    estadosaceptados[0] = 'TOTAL'
                    data['becasestadosaceptados'] = estadosaceptados.items()
                    data['totalesbecasestadosaceptados'] = [becassolicitudes.filter(becaaceptada__in=estadosaceptados.keys() if key == 0 else [key], status=True).count()for key, value in data['becasestadosaceptados']]
                    estadosasignadas = dict(ESTADO_ASIGNACION_BECA)
                    estadosasignadas[0] = 'TOTAL'
                    data['becasestadosasignadas'] = estadosasignadas.items()


                    data['totalesbecasestadosasignadas'] = [becassolicitudes.filter(becaasignada__in=estadosasignadas.keys() if  key == 0 else [key], status=True).count() for key, value in data['becasestadosasignadas']]

                    data['becastipossolicitudes'] = BecaTipo.objects.filter(id__in=becastipos_ids, status=True)
                    becasatipossignaciones_ids = becasasignaciones.values_list('solicitud__becatipo_id', flat=True).distinct()
                    data['becastiposasignaciones'] = BecaTipo.objects.filter(id__in=becasatipossignaciones_ids, status=True)
                    data['total_becassolicitudes'] = becassolicitudes.count()
                    data['total_becasasignaciones'] = becasasignaciones.count()
                    data['recorridos'] = recorridos = BecaSolicitudRecorrido.objects.filter(solicitud_id__in=becassolicitudes_ids, status=True).distinct('estado')

                    return render(request, "estadisticas/tablasegmentobienestaradjudicacadosbecas.html", data)
                except Exception as ex:
                    pass

            elif action == 'segmentobienestarresumenbeca':
                try:
                    data['title'] = u'Resumen de becas'
                    #peridos preinscripcion becas
                    idspb = PreInscripcionBeca.objects.values_list('periodo_id', flat=True).filter(status=True).order_by('periodo_id').distinct('periodo_id')
                    #periodos BecaAsignada
                    idsba = BecaAsignacion.objects.values_list('solicitud__periodo_id', flat=True).filter(status=True).order_by('solicitud__periodo_id').distinct('solicitud__periodo_id')
                    ipsba = idsba.exclude(solicitud__periodo_id__in=idspb)
                    #idps = ipsba | idspb
                    ePeridosBeca = Periodo.objects.filter(Q(pk__in=idspb) | Q(pk__in=ipsba))
                    data['ePeriodosBeca'] = ePeridosBeca
                    data['eBecasResumenes'] = BecaPeriodoResumen.objects.filter(status=True)

                    return render(request, "estadisticas/tablasegmentobienestarresumenbecas.html", data)
                except Exception as ex:
                    pass

            elif action == 'reportematriculadosregularesbienestar':
                try:
                    becaperiodoresumen = BecaPeriodoResumen.objects.filter(pk=int(encrypt(request.GET['idpbr']))).first()
                    if becaperiodoresumen is None:
                        raise NameError("No existe resumen para el periodo seleccionado")
                    __author__ = 'Unemi'

                    fuentenormal = easyxtitulo = easyxf(
                        'font: name Times New Roman, color-index black, bold on , height 350; alignment: horiz centre')
                    titulo = easyxf(
                        'font: name Times New Roman, color-index black, bold on , height 350; alignment: horiz centre')
                    titulo2 = easyxf(
                        'font: name Times New Roman, color-index black, bold on , height 250; alignment: horiz centre')
                    fuentecabecera = easyxf(
                        'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
                    fuentenormal = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
                    fuentemoneda = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right',
                        num_format_str=' "$" #,##0.00')
                    fuentefecha = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz center',
                        num_format_str='yyyy-mm-dd')
                    fuentenumerodecimal = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right',
                        num_format_str='#,##0.00')
                    fuentenumeroentero = easyxf(
                        'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right')

                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('Matriculados')
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = f'attachment; filename=matricualados_regulares: {becaperiodoresumen.periodo}' + random.randint(1,10000).__str__() + '.xls'

                    ws.write_merge(0, 0, 0, 15, 'UNIVERSIDAD ESTATAL DE MILAGRO', titulo)
                    ws.write_merge(1, 1, 0, 15, f'LISTADO DE MATRICULADOS REGULARES DEL PERIODO: {becaperiodoresumen.periodo}', titulo2)

                    row_num = 4
                    columns = [
                        (u"NOMBRES COMPLETOS", 10000),
                        (u"TIPO DE DOCUMENTO", 5000),
                        (u"DOCUMENTO", 5000),
                        (u"FACULTAD", 10000),
                        (u"CARRERA", 15000),
                        (u"PROVINCIA", 5000),
                        (u"CANTÓN", 5000),
                        (u"PARROQUIA", 5000),
                        (u"DIRECCIÓN", 15000),
                        (u"REFERENCIA", 15000),
                        (u"SECTOR", 15000),
                        (u"# CASA", 5000),
                        (u"CORREO PERSONAL", 8000),
                        (u"CORREO UNEMI", 8000),
                        (u"TELÉFONO", 5000),
                        (u"CELULAR", 5000),
                        (u"OPERADORA", 5000),
                        # (u"OBSERVACIÓN", 20000)
                    ]
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], fuentecabecera)
                        ws.col(col_num).width = columns[col_num][1]

                    mallas_ingles = Malla.objects.filter(pk__in=[353, 22]).values_list('id', flat=True)
                    matriculados = Matricula.objects.filter(status=True,
                                                            inscripcion__coordinacion_id__in=[1, 2, 3, 4, 5],
                                                            nivel__periodo=becaperiodoresumen.periodo)\
                                                    .annotate(total_ingles=Count('materiaasignada__materia__asignaturamalla__malla_id',
                                                              filter=Q(materiaasignada__materia__asignaturamalla__malla_id__in=mallas_ingles,
                                                              nivel__periodo=becaperiodoresumen.periodo, status=True)
                                                              ),
                                                              total_general=Count('materiaasignada__materia__asignaturamalla__malla_id',
                                                              filter=Q(nivel__periodo=becaperiodoresumen.periodo, status=True)
                                                              )).exclude(
                                                                        Q(total_general=F('total_ingles')) &
                                                                        ~Q(total_general=0) &
                                                                        ~Q(total_ingles=0)).distinct()#.order_by('apellido1', 'apellido2', 'nombres')
                    matriculados = Matricula.objects.filter(pk__in=matriculados.values_list('id', flat=True))
                    matriculados_regulares = matriculados.filter(matriculagruposocioeconomico__tipomatricula=1,
                                                                 matriculagruposocioeconomico__status=True,
                                                                 retiradomatricula=False)
                    if matriculados_regulares.values_list('id', flat=True).count() >= becaperiodoresumen.matriculados_regulares:
                        matriculados = matriculados_regulares[:becaperiodoresumen.matriculados_regulares]
                    else:
                        matriculados = matriculados.filter(pk__in=matriculados_regulares.values_list('id', flat=True)) | matriculados.exclude(pk__in=matriculados_regulares.values_list('id', flat=True))
                        if matriculados.count() >= becaperiodoresumen.matriculados_regulares:
                            matriculados = matriculados[:becaperiodoresumen.matriculados_regulares]
                        else:
                            maticulasextras = Matricula.objects.filter(status=True,
                                                                       inscripcion__coordinacion_id__in=[1, 2, 3, 4, 5],
                                                                       nivel__periodo=becaperiodoresumen.periodo).exclude(pk__in=matriculados.values_list('id', flat=True))
                            idmls = [id for id in matriculados.values_list('id', flat=True)]
                            idmls.extend([id for id in maticulasextras.values_list('id', flat=True)])
                            matriculados = Matricula.objects.filter(status=True, pk__in=idmls)
                            matriculados = matriculados[:becaperiodoresumen.matriculados_regulares]

                    for matriculado in matriculados:
                        row_num += 1
                        ePersona = matriculado.inscripcion.persona
                        eInscripcion = matriculado.inscripcion
                        ws.write(row_num, 0, ePersona.nombre_completo_inverso(), fuentenormal)
                        ws.write(row_num, 1, ePersona.tipo_documento(), fuentenormal)
                        ws.write(row_num, 2, ePersona.documento(), fuentenormal)
                        ws.write(row_num, 3, eInscripcion.coordinacion.__str__(), fuentenormal)
                        ws.write(row_num, 4, eInscripcion.carrera.__str__(), fuentenormal)
                        ws.write(row_num, 5, str(ePersona.provincia) if ePersona.provincia else '', fuentenormal)
                        ws.write(row_num, 6, str(ePersona.canton) if ePersona.canton else '', fuentenormal)
                        ws.write(row_num, 7, str(ePersona.parroquia) if ePersona.parroquia else '', fuentenormal)
                        ws.write(row_num, 8, ePersona.direccion_corta().upper(), fuentenormal)
                        ws.write(row_num, 9, ePersona.referencia.upper(), fuentenormal)
                        ws.write(row_num, 10, ePersona.sector.upper(), fuentenormal)
                        ws.write(row_num, 11, ePersona.num_direccion, fuentenormal)
                        ws.write(row_num, 12, ePersona.email, fuentenormal)
                        ws.write(row_num, 13, ePersona.emailinst, fuentenormal)
                        ws.write(row_num, 14, ePersona.telefono_conv, fuentenormal)
                        ws.write(row_num, 15, ePersona.telefono, fuentenormal)
                        ws.write(row_num, 16, ePersona.get_tipocelular_display() if ePersona.tipocelular else '', fuentenormal)
                        # cuentabeneficiario = ePersona.cuentabancaria()
                        # ws.write(row_num, 15, cuentabeneficiario.observacion, fuentenormal)

                    wb.save(response)
                    return response
                except Exception as ex:
                    print("Error...")
                    pass

            elif action == 'recaudacionmatriculaspregrado':
                try:
                    from sagest.models import Pago
                    data['title'] = u'Proyección de recaudación matriculas pregrado'
                    data['periodos_pregrado'] = periodos_pregrado = Periodo.objects.filter(status=True, tipo__id=2, id__in=Matricula.objects.values('nivel__periodo__id').filter(nivel__periodo__isnull=False, status=True, rubro__isnull=False).values_list('nivel__periodo__id', flat=True)).distinct().order_by('-fin')
                    data['coordinaciones'] = Coordinacion.objects.filter(id__in=[1,2,3,4,5])
                    data['modalidades'] = Modalidad.objects.filter(status=True)
                    data['nivelesdemallas'] = NivelMalla.objects.all().order_by('id')
                    data['periodosel'] = periodosel = int(request.GET.get('periodosel', 0))
                    facultad, carrera, nivel, modalidad, url_vars=int(request.GET.get('facultad', 0)),\
                                                                   int(request.GET.get('carrera', 0)),\
                                                                   int(request.GET.get('nivel', 0)),\
                                                                   int(request.GET.get('modalidad', 0)), ''
                    if periodosel:
                        data['periodo_'] = periodo_ = Periodo.objects.get(id=periodosel)
                        url_vars += "&periodosel={}".format(periodosel)
                        filtro=Q(status=True, matricula__nivel__periodo=periodo_)
                        filtrop=Q(status=True, rubro__matricula__nivel__periodo=periodo_)
                        if facultad !=0:
                            data['facultad']=facultad
                            filtro= filtro & Q(matricula__inscripcion__coordinacion_id=facultad)
                            filtrop= filtrop & Q(rubro__matricula__inscripcion__coordinacion_id=facultad)
                            url_vars += "&facultad={}".format(facultad)
                        if modalidad !=0:
                            data['modalidad']=modalidad
                            filtro= filtro & Q(matricula__inscripcion__modalidad_id=modalidad)
                            filtrop= filtrop & Q(rubro__matricula__inscripcion__modalidad_id=modalidad)
                            url_vars += "&modalidad={}".format(modalidad)

                        if carrera !=0:
                            data['carrera']=carrera
                            filtro= filtro & Q(matricula__inscripcion__carrera_id=carrera)
                            filtrop= filtrop & Q(rubro__matricula__inscripcion__carrera_id=carrera)
                            url_vars += "&carrera={}".format(carrera)

                        if nivel !=0:
                            data['nivel']=nivel
                            filtro= filtro & Q(matricula__nivelmalla_id=nivel)
                            filtrop= filtrop & Q(rubro__matricula__nivel_id=nivel)
                            url_vars += "&nivel={}".format(nivel)

                        qsrubros = Rubro.objects.filter(filtro)
                        if qsrubros:
                            listado_meses_rubros = qsrubros.values('fechavence__year', 'fechavence__month').annotate(totaladeudado=Coalesce(Sum('valor'), Decimal('0')), totalsaldo=Coalesce(Sum('saldo'), Decimal('0')), totalpagado=Coalesce(Sum('pago__valortotal', filter=Q(pago__status=True)), Decimal('0')), saldocalculado=Coalesce(F('totaladeudado')-F('totalpagado'), Decimal('0'))).order_by('fechavence__year', 'fechavence__month')
                            data['listado_meses_rubros'] = listado_meses_rubros
                            data['totalrubros'] = totalrubros = qsrubros.aggregate(totalrubros=Sum(F('valor'), output_field=FloatField())).get('totalrubros')
                            data['totalpagado'] = totalpagado = qsrubros.aggregate(totalpagado=Sum(F('pago__valortotal'), filter=Q(pago__status=True), output_field=FloatField())).get('totalpagado')
                            data['totalsaldo'] = totalsaldo = (totalrubros - totalpagado)
                            qspagos = Pago.objects.filter(filtrop)
                            listado_pagos_recibidos = qspagos.values('fecha__year', 'fecha__month').annotate(totalpagado=Sum('valortotal')).order_by('fecha__year', 'fecha__month')
                            data['listado_pagos_recibidos'] = listado_pagos_recibidos
                            data['totalpagadopagos'] = totalpagadopagos = qspagos.aggregate(totalpagado=Sum(F('valortotal'), output_field=FloatField())).get('totalpagado')
                            data['total']=len(qsrubros)
                            data['url_vars']=url_vars
                            if 'exportar_excel' in request.GET:
                                wb = openxl.Workbook()
                                wb["Sheet"].title = "Recaudacion_Pregrado"
                                ws = wb.active
                                style_title = openxlFont(name='Arial', size=16, bold=True)
                                style_title2 = openxlFont(name='Arial', size=12, bold=True)
                                style_cab = openxlFont(name='Arial', size=10, bold=True)
                                alinear = alin(horizontal="center", vertical="center")
                                response = HttpResponse(content_type="application/ms-excel")
                                response[
                                    'Content-Disposition'] = 'attachment; filename=Reporte de recaudación matriculas pregrado' + '-' + random.randint(
                                    1, 10000).__str__() + '.xlsx'
                                ws.column_dimensions['B'].width = 25
                                ws.column_dimensions['C'].width = 15
                                ws.column_dimensions['D'].width = 20
                                ws.column_dimensions['E'].width = 10
                                ws.column_dimensions['F'].width = 25
                                ws.column_dimensions['G'].width = 20
                                ws.merge_cells('A1:P1')
                                ws['A1'] = 'REPORTE DE PROYECCIÓN DE RECAUDACIÓN MATRICULAS PREGRADO'
                                celda1 = ws['A1']
                                celda1.font = style_title
                                celda1.alignment = alinear

                                columns = [u"N°", u"NOMBRES Y APELLIDOS", u"CÉDULA",'CORREO', u"TELÉFONO",
                                           u"FACULTAD", u"CARRERA", u"MODALIDAD", u"NIVEL",u"NOMBRE RUBRO",
                                           u"VALOR",u"VALOR IVA",u"VALOR TOTAL","CANCELADO"
                                           ]
                                row_num = 3
                                for col_num in range(0, len(columns)):
                                    celda = ws.cell(row=row_num, column=(col_num + 1), value=columns[col_num])
                                    celda.font = style_cab

                                mensaje = 'NO REGISTRA'
                                row_num = 4
                                numero = 0
                                for rubro in qsrubros:
                                    numero += 1
                                    ws.cell(row=row_num, column=1, value=numero)
                                    ws.cell(row=row_num, column=2, value=str(rubro.persona))
                                    ws.cell(row=row_num, column=3, value=str(rubro.persona.cedula))
                                    ws.cell(row=row_num, column=4, value=str(rubro.persona.email))
                                    ws.cell(row=row_num, column=5, value=str(rubro.persona.telefono))
                                    ws.cell(row=row_num, column=6, value=str(rubro.matricula.inscripcion.coordinacion))
                                    ws.cell(row=row_num, column=7, value=str(rubro.matricula.inscripcion.carrera))
                                    ws.cell(row=row_num, column=8, value=str(rubro.matricula.inscripcion.modalidad))
                                    ws.cell(row=row_num, column=9, value=str(rubro.matricula.nivelmalla))
                                    ws.cell(row=row_num, column=10, value=str(rubro.nombre))
                                    ws.cell(row=row_num, column=11, value=f'${rubro.valor}')
                                    ws.cell(row=row_num, column=12, value=f'${rubro.valoriva}')
                                    ws.cell(row=row_num, column=13, value=f'${rubro.valortotal}')
                                    ws.cell(row=row_num, column=14, value='SI'if rubro.cancelado else 'NO')
                                    row_num += 1

                                ws2 = wb.create_sheet("Proyeccion_Mensual")
                                ws2.column_dimensions['B'].width = 20
                                ws2.column_dimensions['C'].width = 20
                                ws2.column_dimensions['D'].width = 25
                                ws2.column_dimensions['E'].width = 25
                                ws2.merge_cells('A1:E1')
                                ws2['A1'] = 'PROYECCION DE PERIODO {}'.format(periodo_)
                                celdaA1 = ws2['A1']
                                celdaA1.font = style_title2
                                celdaA1.alignment = alinear
                                columns = [u"N°", u"FECHA", u"TOTAL RUBROS", 'TOTAL PAGADO', u"PENDIENTE"]
                                row_num = 2
                                for col_num in range(0, len(columns)):
                                    celda = ws2.cell(row=row_num, column=(col_num + 1), value=columns[col_num])
                                    celda.font = style_cab

                                row_num = 3
                                numero = 0
                                for l in listado_meses_rubros:
                                    numero += 1
                                    ws2.cell(row=row_num, column=1, value=numero)
                                    ws2.cell(row=row_num, column=2, value=f'{l["fechavence__month"]}/{l["fechavence__year"]}')
                                    ws2.cell(row=row_num, column=3, value=f'${l["totaladeudado"]}')
                                    ws2.cell(row=row_num, column=4, value=f'${l["totalpagado"]}')
                                    ws2.cell(row=row_num, column=5, value=f'${l["saldocalculado"]}')
                                    row_num += 1
                                ws2.merge_cells(f'A{row_num}:B{row_num}')
                                ws2[f'A{row_num}'] = 'TOTAL'
                                ws2.cell(row=row_num, column=3, value=f'${totalrubros}')
                                ws2.cell(row=row_num, column=4, value=f'${totalpagado}')
                                ws2.cell(row=row_num, column=5, value=f'${totalsaldo}')
                                wb.save(response)
                                return response
                    return render(request, "estadisticas/recaudacionpregrado/view.html", data)
                except Exception as ex:
                    pass

            elif action == 'listcarreras':
                try:
                    idc, idm=int(request.GET.get('id',0)), int(request.GET.get('idm',0))
                    if idc > 0:
                        coordinacion = Coordinacion.objects.get(pk=request.GET['id'])
                        carreras=coordinacion.carreras()
                        if idm > 0:
                           carreras=carreras.filter(modalidad=idm)
                    elif idm > 0:
                        carreras = Carrera.objects.filter(modalidad=idm)
                    lista = []
                    for carrera in carreras:
                        lista.append([carrera.id, carrera.nombre])
                    return JsonResponse({'result': True, 'lista': lista})
                except Exception as ex:
                    return JsonResponse({"result": False, "mensaje": u"Error al obtener los datos."})

            elif action == 'segmentoencuantasatifacciongeneral_old':
                try:
                    idprocesos = RespuestaEncuestaSatisfaccion.objects.filter(status=True).values_list('pregunta__encuesta__proceso_id', flat=True)
                    data['procesos'] = procesos = Proceso.objects.filter(id__in=idprocesos)
                    proceso_id = procesos.first().id if procesos.first() else None
                    if 'idproc' in request.GET:
                        proceso_id = int(encrypt(request.GET['idproc']))
                    data['proceso_id'] = proceso_id
                    data['eProcesos'] = procesos.filter(id=proceso_id)
                    return render(request, "estadisticas/tablasegmentoencuantasatifacciongeneral.html", data)
                except Exception as ex:
                    pass

            elif action == 'segmentoencuantasatifacciongeneral_reporte_excel_old':
                try:
                    categoria = CategoriaEncuesta.objects.get(id=encrypt_id(request.GET['idcategoria']))
                    idprocesos = RespuestaEncuestaSatisfaccion.objects.filter(status=True).values_list('pregunta__encuesta__proceso_id', flat=True)
                    data['procesos'] = eProcesos = Proceso.objects.filter(id__in=idprocesos)
                    if 'idproc' in request.GET:
                        proceso_id = int(encrypt(request.GET['idproc']))
                        eProcesos = eProcesos.filter(id=proceso_id)

                    output = io.BytesIO()
                    __author__ = 'Unemi'
                    workbook = xlsxwriter.Workbook(output, {'constant_memory': True})
                    row_last = 0
                    for eProceso in eProcesos:
                        encuesta_proceso = eProceso.encuesta_proceso()
                        name_worksheet = eProceso.sigla.strip().replace(' ', '-')
                        worksheet = workbook.add_worksheet(name_worksheet)
                        fuentecabecera = workbook.add_format({
                            'align': 'center',
                            'bg_color': 'silver',
                            'border': 1,
                            'bold': 1
                        })

                        formatoceldacenter = workbook.add_format({
                            'border': 1,
                            'valign': 'vcenter',
                            'align': 'center'})

                        worksheet.merge_range(0, 0, 0, encuesta_proceso.valoracion_colspan(), eProceso.descripcion, fuentecabecera)
                        worksheet.merge_range(1, 1, 1, encuesta_proceso.valoracion_colspan(), 'Nivel de satisfacción',formatoceldacenter)
                        worksheet.merge_range(1, 0, 2, 0, 'Criterios', formatoceldacenter)
                        worksheet.set_column(0, 0, 40)
                        col_temp = 0
                        listado_estrellas = encuesta_proceso.lista_valoracion()

                        for index, valor in enumerate(listado_estrellas):
                            col_temp += 1
                            worksheet.write(2, col_temp, f'{valor} {"⭐"*valor}', formatoceldacenter)
                            worksheet.set_column(2, col_temp, 15)
                        data_categories = [name_worksheet, 2, 1, 2, col_temp]
                        worksheet.write(2, col_temp+1, 'TOTAL', formatoceldacenter)
                        worksheet.set_column(2, col_temp+1, 15)
                        row_num, col_num = 3, 0
                        data_grafico = []
                        x_offset, y_offset = 0,0
                        for pregunta in encuesta_proceso.preguntas_para_estadisticas():
                            col_num = 0
                            worksheet.write(row_num, col_num, pregunta.descripcion, formatoceldacenter)
                            dict_data = {
                                'title': pregunta.descripcion,
                                'col_num_first': col_num + 1,
                                'row_num_first': row_num,
                                'row_num_last': row_num,
                            }
                            for valor in listado_estrellas:
                                col_num += 1
                                cantidad = pregunta.cantidad_valoracion_respuesta_encuesta_estadistica(valor)
                                worksheet.write(row_num, col_num, cantidad, formatoceldacenter)
                            dict_data['col_num_last'] = col_num
                            x_offset += 25
                            y_offset += 10
                            dict_data['x_offset'] = x_offset
                            dict_data['y_offset'] = y_offset
                            col_num += 1
                            total = pregunta.cantidad_total_respuesta_encuesta_estadistica()
                            worksheet.write(row_num, col_num, total, formatoceldacenter)
                            row_num += 1
                            data_grafico.append(dict_data)
                            row_last = row_num

                        #worksheet.merge_range(1, 2, 1, encuesta_proceso.valoracion_colspan_general(), 'Criterios', formatoceldacenter)
                        row_num, numcolum = 0, 0
                        col_index = 1
                        for dg in data_grafico:
                            chart3 = workbook.add_chart({'type': 'pie'})

                            # Configure the series.
                            chart3.add_series({
                                'name': u'%s' % dg['title'],
                                'categories': data_categories,
                                'values': [name_worksheet, dg['row_num_first'], dg['col_num_first'], dg['row_num_last'], dg['col_num_last']],
                                'data_labels': {'percentage': True},
                            })

                            # Add a title.
                            chart3.set_title({'name':u'%s' % dg['title']})

                            # Change the angle/rotation of the first segment.
                            chart3.set_rotation(90)
                            chart3.set_size({'width': 324, 'height': 288})
                            # Insert the chart into the worksheet (with an offset).
                            worksheet.insert_chart(row_last + 2, col_index, chart3, {'x_offset': 7, 'y_offset': 10})
                            col_index += 3

                    workbook.close()
                    output.seek(0)
                    name_document = 'reporte_encuesta_satisfacion{}'.format(str(datetime.now().strftime('%Y%m%d_%H%M%S')))
                    filename = f'{name_document}.xlsx'
                    response = HttpResponse(
                        output,
                        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                    )
                    response['Content-Disposition'] = 'attachment; filename=%s' % filename

                    return response
                except Exception as ex:
                    pass

            elif action == 'segmentoencuantasatifaccion_reporte_excel_resultados':
                try:
                    categoria = CategoriaEncuesta.objects.get(id=encrypt_id(request.GET['idcategoria']))
                    # idprocesos = RespuestaEncuestaSatisfaccion.objects.filter(status=True, pergunta__encuesta__categoria=categoria).values_list('pregunta__encuesta_id', flat=True)
                    data['procesos'] = encuestas = categoria.encuestas()
                    idencuesta = request.GET.get('idencuesta', '')
                    if idencuesta:
                        encuestas = encuestas.filter(id=idencuesta)

                    output = io.BytesIO()
                    __author__ = 'Unemi'
                    workbook = xlsxwriter.Workbook(output, {'constant_memory': True})
                    row_last = 0
                    for idex, encuesta in enumerate(encuestas, start=1):
                        encuesta_objeto = encuesta.encuesta_objeto()
                        name_worksheet = f'{idex}.'+encuesta.sigla_text()
                        worksheet = workbook.add_worksheet(name_worksheet)
                        fuentecabecera = workbook.add_format({
                            'align': 'center',
                            'bg_color': 'silver',
                            'border': 1,
                            'bold': 1
                        })

                        formatoceldacenter = workbook.add_format({
                            'border': 1,
                            'valign': 'vcenter',
                            'bold': 1,
                            'align': 'center'})

                        formatocenter = workbook.add_format({
                            'valign': 'vcenter',
                            'align': 'center'})

                        columns = [
                            (u"Solicitante", 40),
                            (u"Tipo de usuario", 15),
                            (u"Cargo", 50),
                            (u"Cédula", 15),
                            (u"Correo", 30),
                            (u"Requerimiento", 75),
                            (u"Fecha", 15),
                            (u"Persona asignada", 40),
                            (u"Observación", 30),
                        ]

                        rowtitlewidth = len(columns) + encuesta.preguntas_para_estadisticas().count() - 1
                        worksheet.merge_range(0, 0, 0, rowtitlewidth, encuesta_objeto.name_estadistica(), fuentecabecera)

                        for index, pregunta in enumerate(encuesta.preguntas_para_estadisticas(), start=1):
                            columns.append((f'{index}. {pregunta.descripcion}', 60))

                        row_num = 2
                        for col_num, (header, width) in enumerate(columns):
                            worksheet.write(row_num, col_num, header, formatoceldacenter)
                            worksheet.set_column(col_num, col_num, width)

                        incidencias = encuesta_objeto.reservas()

                        for index, incidencia in enumerate(incidencias, start=3):
                            datosreporte = incidencia.get_datos_reporte_encuesta_satisfaccion()
                            respuestasencuestas = incidencia.respuestas_encuesta()
                            if respuestasencuestas:
                                row_num += 1
                                worksheet.write(row_num, 0, datosreporte['solicitante'], formatocenter)
                                worksheet.write(row_num, 1, datosreporte['tipousuario'], formatocenter)
                                worksheet.write(row_num, 2, datosreporte['cargo'], formatocenter)
                                worksheet.write(row_num, 3, datosreporte['cedula'], formatocenter)
                                worksheet.write(row_num, 4, datosreporte['correo'], formatocenter)
                                worksheet.write(row_num, 5, datosreporte['requerimiento'], formatocenter)
                                worksheet.write(row_num, 6, datosreporte['fecha'], formatocenter)
                                worksheet.write(row_num, 7, datosreporte['asiganadoa'], formatocenter)
                                worksheet.write(row_num, 8, respuestasencuestas[0].observacion, formatocenter)

                                for indexcolumn, respuesta in enumerate(respuestasencuestas, start=9):
                                    worksheet.write(row_num, indexcolumn, f'{respuesta.valoracion}', formatocenter)

                    workbook.close()
                    output.seek(0)
                    name_document = 'reporte_encuesta_satisfacion_resultados{}'.format(str(datetime.now().strftime('%Y%m%d_%H%M%S')))
                    filename = f'{name_document}.xlsx'
                    response = HttpResponse(
                        output,
                        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                    )
                    response['Content-Disposition'] = 'attachment; filename=%s' % filename

                    return response
                except Exception as ex:
                    pass

            elif action == 'segmentoencuantasatifacciongeneral_reporte_excel':
                try:
                    categoria = CategoriaEncuesta.objects.get(id=encrypt_id(request.GET['idcategoria']))
                    # idprocesos = RespuestaEncuestaSatisfaccion.objects.filter(status=True, pergunta__encuesta__categoria=categoria).values_list('pregunta__encuesta_id', flat=True)
                    data['procesos'] = encuestas = categoria.encuestas()
                    idencuesta = request.GET.get('idencuesta', '')
                    if idencuesta:
                        encuestas = encuestas.filter(id=idencuesta)

                    output = io.BytesIO()
                    __author__ = 'Unemi'
                    workbook = xlsxwriter.Workbook(output, {'constant_memory': True})
                    row_last = 0
                    for idex, encuesta in enumerate(encuestas, start=1):
                        encuesta_objeto = encuesta.encuesta_objeto()
                        name_worksheet = f'{idex}.'+encuesta.sigla_text()
                        worksheet = workbook.add_worksheet(name_worksheet)
                        fuentecabecera = workbook.add_format({
                            'align': 'center',
                            'bg_color': 'silver',
                            'border': 1,
                            'bold': 1
                        })

                        formatoceldacenter = workbook.add_format({
                            'border': 1,
                            'valign': 'vcenter',
                            'align': 'center'})

                        worksheet.merge_range(0, 0, 0, encuesta.valoracion_colspan(), encuesta_objeto.name_estadistica(), fuentecabecera)
                        worksheet.merge_range(1, 1, 1, encuesta.valoracion_colspan(), 'Nivel de satisfacción',formatoceldacenter)
                        worksheet.merge_range(1, 0, 2, 0, 'Criterios', formatoceldacenter)
                        worksheet.set_column(0, 0, 40)
                        col_temp = 0
                        listado_estrellas = encuesta.lista_valoracion()

                        for index, valor in enumerate(listado_estrellas):
                            col_temp += 1
                            worksheet.write(2, col_temp, f'{valor} {"⭐"*valor}', formatoceldacenter)
                            worksheet.set_column(2, col_temp, 15)
                        data_categories = [name_worksheet, 2, 1, 2, col_temp]
                        worksheet.write(2, col_temp+1, 'TOTAL', formatoceldacenter)
                        worksheet.set_column(2, col_temp+1, 15)
                        row_num, col_num = 3, 0
                        data_grafico = []
                        x_offset, y_offset = 0,0
                        for pregunta in encuesta.preguntas_para_estadisticas():
                            col_num = 0
                            worksheet.write(row_num, col_num, pregunta.descripcion, formatoceldacenter)
                            dict_data = {
                                'title': pregunta.descripcion,
                                'col_num_first': col_num + 1,
                                'row_num_first': row_num,
                                'row_num_last': row_num,
                            }
                            for valor in listado_estrellas:
                                col_num += 1
                                cantidad = pregunta.cantidad_valoracion_respuesta_encuesta_estadistica(valor)
                                worksheet.write(row_num, col_num, cantidad, formatoceldacenter)
                            dict_data['col_num_last'] = col_num
                            x_offset += 25
                            y_offset += 10
                            dict_data['x_offset'] = x_offset
                            dict_data['y_offset'] = y_offset
                            col_num += 1
                            total = pregunta.cantidad_total_respuesta_encuesta_estadistica()
                            worksheet.write(row_num, col_num, total, formatoceldacenter)
                            row_num += 1
                            data_grafico.append(dict_data)
                            row_last = row_num

                        #worksheet.merge_range(1, 2, 1, encuesta_proceso.valoracion_colspan_general(), 'Criterios', formatoceldacenter)
                        row_num, numcolum = 0, 0
                        col_index = 1
                        for dg in data_grafico:
                            chart3 = workbook.add_chart({'type': 'pie'})

                            # Configure the series.
                            chart3.add_series({
                                'name': u'%s' % dg['title'],
                                'categories': data_categories,
                                'values': [name_worksheet, dg['row_num_first'], dg['col_num_first'], dg['row_num_last'], dg['col_num_last']],
                                'data_labels': {'percentage': True},
                            })

                            # Add a title.
                            chart3.set_title({'name':u'%s' % dg['title']})

                            # Change the angle/rotation of the first segment.
                            chart3.set_rotation(90)
                            chart3.set_size({'width': 324, 'height': 288})
                            # Insert the chart into the worksheet (with an offset).
                            worksheet.insert_chart(row_last + 2, col_index, chart3, {'x_offset': 7, 'y_offset': 10})
                            col_index += 3

                    workbook.close()
                    output.seek(0)
                    name_document = 'reporte_encuesta_satisfacion{}'.format(str(datetime.now().strftime('%Y%m%d_%H%M%S')))
                    filename = f'{name_document}.xlsx'
                    response = HttpResponse(
                        output,
                        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                    )
                    response['Content-Disposition'] = 'attachment; filename=%s' % filename

                    return response
                except Exception as ex:
                    pass

            elif action == 'segmentoencuantasatifacciongeneral':
                try:
                    data['title'] = 'Estadísticas | Encuestas'
                    data['categorias'] = categorias = CategoriaEncuesta.objects.filter(status=True)
                    return render(request, "estadisticas/estadisticasencuesta.html", data)
                except Exception as ex:
                    pass

            elif action == 'cargarestadistica':
                try:
                    datos = {}
                    id_c, id_sc = request.GET.get('idcategoria', ''), request.GET.get('idsubcategoria', '')
                    categoria = CategoriaEncuesta.objects.get(id=encrypt_id(id_c))
                    encuesta = EncuestaProceso.objects.get(id=id_sc)
                    preguntas = encuesta.preguntas_para_estadisticas()
                    for p in preguntas:
                        datos[p.descripcion] ={'id': p.id, 'estadistica':p.valores_estadisticos(encuesta.valoracion)}
                    return JsonResponse({'results': True, 'data': datos, 'valoracion': encuesta.valoracion, 'title': encuesta.name_estadistica()})
                except Exception as ex:
                    pass

            elif action == 'cargarsubcategorias':
                try:
                    context = []
                    categoria_id = encrypt_id(request.GET['id'])
                    encuestas = EncuestaProceso.objects.filter(categoria_id=categoria_id, status=True, content_type__isnull=False).order_by('object_id')
                    iddefault = ''
                    for encuesta in encuestas:
                        if encuesta.preguntas_obj():
                            context.append({'text': encuesta.name_estadistica(), 'value': encuesta.id})
                            if not iddefault:
                                iddefault = encuesta.id
                    return JsonResponse({'results': True, 'data': context, 'iddefault':iddefault})
                except Exception as ex:
                    pass

            elif action == 'matriculadosasistencia':
                try:
                    data['title'] = u'Matriculados por asistencia'
                    data['periodo'] = periodo

                    data['hoy'] = hoy = datetime.now()
                    data['ahora'] = ahora = hoy.time()

                    data['matriculados_coordinacion'] = matriculados_coordinacion = Matricula.objects.filter(status=True, bloqueomatricula=False, retiradomatricula=False, inscripcion__carrera__coordinacion__id__in=[1, 2, 3, 4, 5], nivel__periodo=periodo).exclude(inscripcion__carrera__modalidad=3)
                    lista_materiaasignada = MateriaAsignada.objects.filter(matricula__in=matriculados_coordinacion, status=True)
                    data['indicadores'] = indicadores = [{"id": 1, "nombre": u'ASISTENCIAS DE MATRICULADOS POR FECHA CORTE'},
                                                         {"id": 2, "nombre": u'ASISTENCIAS DE MATRICULADOS POR TURNO CORTE'}]
                    data['jornadas'] = Sesion.objects.filter(pk__in=lista_materiaasignada.values_list('materia__nivel__sesion_id', flat=True).distinct())
                    data['modalidades'] = Modalidad.objects.filter(status=True).exclude(pk=3)

                    genera, _indicador, jornada, modalidad, fecha = int(request.GET.get('genera', 0)), int(request.GET.get('indicador', 1)), int(request.GET.get('jornada', 0)), int(request.GET.get('modalidad', 0)), request.GET.get('fecha', '')

                    filtro, url_vars = Q(status=True), ''
                    fecha = datetime.strptime(fecha, '%Y-%m-%d').date() if fecha else fecha

                    if _indicador > 0:
                        filtro &= Q(indicador=_indicador)
                    data['indicador'] = _indicador

                    if modalidad > 0 and modalidad != 100:
                        filtro &= Q(carrera__modalidad=modalidad)
                    data['modalidad'] = modalidad

                    if jornada > 0 and jornada != 100:
                        filtro &= Q(sesion_id=jornada)
                    data['jornada'] = jornada

                    listado = AsistenciaMatriculadoPeriodo.objects.filter(filtro).order_by('fecha')
                    if not fecha and listado:
                        fecha = listado.last().fecha
                    if fecha: listado = listado.filter(fecha=fecha)
                    data['fecha'] = fecha

                    data['texto_indicador'] = texto_indicador = next((indicador["nombre"] for indicador in indicadores if indicador["id"] == _indicador), None)
                    if genera == 1:
                        if _indicador <= 0 and jornada <= 0 and modalidad <= 0 or not fecha: return JsonResponse({"result": False, "mensaje": u"Debe seleccionar un indicador.."})
                        if jornada <= 0 or modalidad <= 0: return JsonResponse({"result": False, "mensaje": u"Debe seleccionar los campos correctamente.."})
                        if not fecha: return JsonResponse({"result": False, "mensaje": u"Debe seleccionar una fecha.."})

                        noti = Notificacion(cuerpo='Generación de reporte de excel en progreso',
                                            titulo=f'Reporte excel de {texto_indicador}', destinatario=persona, url='', prioridad=1, app_label='SGA',
                                            fecha_hora_visible=datetime.now() + timedelta(days=1), tipo=2, en_proceso=True)
                        noti.save(request)
                        reporte_matriculados_asistencia_background(request=request, data=data, notif=noti.pk, periodo=periodo).start()

                        return JsonResponse({"result": True,
                                             "mensaje": u"El reporte se está realizando. Verifique su apartado de notificaciones después de unos minutos.",
                                             "btn_notificaciones": traerNotificaciones(request, data, persona)})

                    # Gráficos
                    sesiones_data = {}
                    if _indicador == 1:
                        # Diccionario para agrupar datos por sesión
                        sesiones_aggregated = defaultdict(lambda: defaultdict(lambda: {'matriculados': 0, 'asistieron': 0, 'noasistieron': 0}))

                        # Agrupar datos por sesión y luego por carrera
                        for asistencia in listado:
                            sesion_key = asistencia.sesion.nombre
                            carrera_key = asistencia.carrera.__str__().capitalize()
                            sesiones_aggregated[sesion_key][carrera_key]['matriculados'] += asistencia.matriculados
                            sesiones_aggregated[sesion_key][carrera_key]['asistieron'] += asistencia.asistieron
                            sesiones_aggregated[sesion_key][carrera_key]['noasistieron'] += asistencia.noasistieron

                        # Preparar datos para el template
                        for sesion_key, carreras in sesiones_aggregated.items():
                            categories = list(carreras.keys())
                            matriculados_data = [carreras[carrera]['matriculados'] for carrera in categories]
                            asistieron_data = [carreras[carrera]['asistieron'] for carrera in categories]
                            noasistieron_data = [carreras[carrera]['noasistieron'] for carrera in categories]
                            sesiones_data[sesion_key] = {
                                'categories': json.dumps(categories),
                                'matriculados_data': json.dumps(matriculados_data),
                                'asistieron_data': json.dumps(asistieron_data),
                                'noasistieron_data': json.dumps(noasistieron_data),
                            }

                    elif _indicador == 2:
                        # Diccionario para agrupar datos por sesión y luego por turno
                        sesiones_aggregated = defaultdict(lambda: defaultdict(lambda: {'docentesturno': 0, 'matriculados': 0, 'asistieron': 0, 'noasistieron': 0}))

                        # Agrupar datos por sesión y luego por turno
                        for asistencia in listado:
                            sesion_key = asistencia.sesion.nombre
                            turno_key = asistencia.turno.nombre_horario()
                            sesiones_aggregated[sesion_key][turno_key]['docentesturno'] += asistencia.docentesturno
                            sesiones_aggregated[sesion_key][turno_key]['matriculados'] += asistencia.matriculados
                            sesiones_aggregated[sesion_key][turno_key]['asistieron'] += asistencia.asistieron
                            sesiones_aggregated[sesion_key][turno_key]['noasistieron'] += asistencia.noasistieron

                        # Preparar datos para el template
                        for sesion_key, turnos in sesiones_aggregated.items():
                            categories = list(turnos.keys())
                            docentesturno_data = [turnos[turno]['docentesturno'] for turno in categories]
                            matriculados_data = [turnos[turno]['matriculados'] for turno in categories]
                            asistieron_data = [turnos[turno]['asistieron'] for turno in categories]
                            noasistieron_data = [turnos[turno]['noasistieron'] for turno in categories]
                            sesiones_data[sesion_key] = {
                                'categories': json.dumps(categories),
                                'docentesturno_data': json.dumps(docentesturno_data),
                                'matriculados_data': json.dumps(matriculados_data),
                                'asistieron_data': json.dumps(asistieron_data),
                                'noasistieron_data': json.dumps(noasistieron_data),
                            }

                    data['listado'] = listado
                    data['fecha'] = fecha
                    data['sesiones_data'] = sesiones_data
                    data['url_vars'] = url_vars
                    return render(request, "estadisticas/matriculadoasistencia.html", data)
                except Exception as ex:
                    pass

            elif action == 'detallemodulosest':
                try:
                    filtro = Q(status=True, materia__nivel__periodo__tipo__id=3, tipoprofesor__id=11,
                               materia__fin__lt=datetime.now().date())

                    desde = hasta = ''
                    url_vars = ' '
                    ide = idc = None

                    if 'desde' in request.GET:
                        desde = request.GET['desde']
                    if 'hasta' in request.GET:
                        hasta = request.GET['hasta']
                    if 'ide' in request.GET:
                        ide = request.GET['ide']
                    if 'idc' in request.GET:
                        idc = request.GET['idc']

                    if ide:
                        filtro = filtro & Q(materia__asignaturamalla__malla__carrera__escuelaposgrado__id=int(ide))

                    if idc:
                        filtro = filtro & Q(materia__asignaturamalla__malla__carrera__id=int(idc))

                    if desde and hasta:
                        data['desde'] = desde
                        data['hasta'] = hasta
                        filtro = filtro & Q(materia__fin__range=(desde, hasta))
                        url_vars += "&desde={}".format(desde)
                        url_vars += "&hasta={}".format(hasta)

                    elif desde:
                        data['desde'] = desde
                        filtro = filtro & Q(materia__fin__gte=desde)
                        url_vars += "&desde={}".format(hasta)

                    elif hasta:
                        data['hasta'] = hasta
                        filtro = filtro & Q(materia__fin__lte=hasta)
                        url_vars += "&hasta={}".format(hasta)

                    if desde == '' and hasta == '':
                        lista = []
                    else:
                        lista = []
                        eModulos = ProfesorMateria.objects.filter(filtro).order_by('-id')
                        for eModulo in eModulos:
                            eProfesor = eModulo.profesor
                            distributivo = eProfesor.distributivohoraseval(eModulo.materia.nivel.periodo)
                            if distributivo and distributivo.resumen_evaluacion_acreditacion():
                                resumen = distributivo.resumen_evaluacion_acreditacion()
                                if resumen.resultado_docencia and resumen.resultado_docencia > 0:
                                    lista.append(eModulo.id)
                    eModulos = ProfesorMateria.objects.filter(id__in=lista).order_by('-id')
                    data['eModulos'] = eModulos
                    template = get_template("estadisticas/listadocentesmod.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

            return HttpResponseRedirect(request.path)
        else:
            data['title'] = u'Estadisticas y graficos'
            data['hoy'] = datetime.now().date()
            data['hoy_str'] = str(datetime.now().date())
            return render(request, "estadisticas/estadisticas.html", data)

