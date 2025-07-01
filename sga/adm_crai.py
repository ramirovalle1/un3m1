from datetime import datetime, timedelta
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Q
from django.http import HttpResponseRedirect, JsonResponse, HttpResponse
from django.shortcuts import render
from django.forms.models import model_to_dict
from django.template import Context
from django.template.loader import get_template
from openpyxl import Workbook
from xlwt import *
import random
import json
from decorators import secure_module, last_access
from sagest.forms import ClienteExternoCraiForm, CubiculoForm
from sagest.funciones import encrypt_id
from sagest.models import DistributivoPersona, CubiculoCrai, PisosChoice
from sga.commonviews import adduserdata
from sga.excelbackground import reporte_acceso_crai_background
from sga.forms import RegistrarIngresoCraiForm, RegistroEstudianteForm, RegistroDocenteAdministrativoForm, \
    RegistroExternoForm
from sga.funciones import MiPaginador, log, convertir_fecha
from sga.funcionesxhtml2pdf import conviert_html_to_pdf
from sga.models import Inscripcion, RegistrarIngresoCrai, RegistrarIngresoCrai, Externo, \
    Persona, ActividadesCrai, RegistrarActividadesCrai, TipoServicioCrai, Profesor, LibroKohaProgramaAnaliticoAsignatura, Notificacion
from sga.templatetags.sga_extras import encrypt

@login_required(redirect_field_name='ret', login_url='/loginsga')
@secure_module
@last_access
@transaction.atomic()

def view(request):
    data = {}
    adduserdata(request, data)
    persona = request.session['persona']
    data['periodo'] = periodo = request.session['periodo']
    if request.method == 'POST':
        action = request.POST['action']
        if action == 'addinscripcion':
            try:
                if 'id' in request.POST:
                    form = RegistrarIngresoCraiForm(request.POST)
                    inscripcion = Inscripcion.objects.get(pk=int(request.POST['id']))
                    if  'profesor' in request.POST and 'tiposervicio' in request.POST :
                        reg = RegistrarIngresoCrai(inscripcion=inscripcion,
                                                   persona_id=inscripcion.persona.pk,
                                                   profesor_id=request.POST['profesor'],
                                                   tiposerviciocrai_id= request.POST['tiposervicio'],
                                                   actividad= request.POST['actividad'],
                                                   fecha=datetime.now().date(),
                                                   horainicio=datetime.now().time())
                    if  'tiposerv' in request.POST:
                        reg = RegistrarIngresoCrai(inscripcion=inscripcion,
                                                   tiposerviciocrai_id=request.POST['tiposerv'],
                                                   fecha=datetime.now().date(),
                                                   horainicio=datetime.now().time())
                    reg.save(request)
                    log(u'Adicionó una nueva visita de incripcion crai: %s' % reg.inscripcion.persona, request, "add")
                    return JsonResponse({"result": "ok", "mensaje": u'Se registro correctamente...'})
                else:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        if action == 'add_docenteadministrativo':
            try:
                if 'id' in request.POST:
                    administrativo = DistributivoPersona.objects.get(pk=int(request.POST['id']))
                    reg = RegistrarIngresoCrai(persona=administrativo.persona,
                                               tiposerviciocrai_id=1,
                                               actividad='BIBLIOTECA',
                                               regimenlaboral=administrativo.regimenlaboral,
                                               fecha=datetime.now().date(),
                                               horainicio=datetime.now().time())
                    reg.save(request)
                    log(u'Adicionó una nueva visita de: %s' % reg.persona, request,"add")
                    return JsonResponse({"result": "ok", "mensaje": u'Se registro correctamente...'})
                else:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        if action == 'addinscripcionexterno':
            try:
                if 'id' in request.POST:
                    form = RegistrarIngresoCraiForm(request.POST)
                    if form.is_valid():
                        persona = Persona.objects.get(pk=int(request.POST['id']))
                        reg = RegistrarIngresoCrai(persona=persona,
                                                   profesor_id=form.cleaned_data['profesor'],
                                                   tiposerviciocrai=form.cleaned_data['tiposerviciocrai'],
                                                   actividad=form.cleaned_data['actividad'],
                                                   fecha=datetime.now().date(),
                                                   horainicio=datetime.now().time())
                        reg.save(request)
                        log(u'Adicionó una nueva visita de incripcion externo crai: %s' % reg.persona, request, "add")
                        return JsonResponse({"result": "ok", "mensaje": u'Se registro correctamente...'})
                else:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        if action == 'addinscripcionbiblioteca':
            try:
                persona = Persona.objects.get(inscripcion=int(request.GET['id'])).pk
                inscripcion = Inscripcion.objects.get(pk=int(request.POST['id']))
                reg = RegistrarIngresoCrai(inscripcion=inscripcion,
                                           persona_id=persona,
                                           tiposerviciocrai_id=1,
                                           actividad='BIBLIOTECA',
                                           fecha=datetime.now().date(),
                                           horainicio=datetime.now().time())
                reg.save(request)
                log(u'Adicionó una nueva visita de incripcion crai: %s' % reg.inscripcion.persona, request, "add")
                return JsonResponse({"result": "ok", "mensaje": u'Se registro correctamente...'})

            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al enviar los datos."})


        if action == 'addaux':
            try:
                persona = Persona.objects.get(inscripcion=int(request.GET['id'])).pk
                inscripcion = Inscripcion.objects.get(pk=int(request.POST['id']))
                if not RegistrarIngresoCrai.objects.filter(inscripcion=inscripcion, tiposerviciocrai_id=1, actividad='BIBLIOTECA', fecha=datetime.now().date()).exists():
                    reg = RegistrarIngresoCrai(inscripcion=inscripcion,
                                               persona_id=persona,
                                               tiposerviciocrai_id=1,
                                               actividad='BIBLIOTECA',
                                               fecha=datetime.now().date(),
                                               horainicio=datetime.now().time())
                    reg.save(request)
                    log(u'Adicionó una nueva visita de incripcion crai: %s' % reg.inscripcion.persona, request, "add")
                else:
                    regaux = RegistrarIngresoCrai.objects.filter(inscripcion=inscripcion, tiposerviciocrai_id=1, actividad='BIBLIOTECA', fecha=datetime.now().date())[0]
                    horaingreso = datetime(regaux.fecha.year, regaux.fecha.month, regaux.fecha.day, regaux.horainicio.hour, regaux.horainicio.minute, regaux.horainicio.second) + timedelta(hours=3, minutes=00, seconds=00)
                    horaconsulta = datetime.now()
                    if horaconsulta > horaingreso:
                        reg = RegistrarIngresoCrai(inscripcion=inscripcion,
                                                   persona_id=persona,
                                                   tiposerviciocrai_id=1,
                                                   actividad='BIBLIOTECA',
                                                   fecha=datetime.now().date(),
                                                   horainicio=datetime.now().time())
                        reg.save(request)
                        log(u'Adicionó una nueva visita de incripcion crai: %s' % reg.inscripcion.persona, request, "add")
                    else:
                        reg = regaux

                if RegistrarActividadesCrai.objects.filter(registraringresocrai=reg, actividadescrai_id=int(request.POST['ida'])).exists():
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": "Actividad ya seleccionada."})
                reg1 = RegistrarActividadesCrai(registraringresocrai=reg, actividadescrai_id=int(request.POST['ida']))
                reg1.save(request)
                log(u'Adicionó una actividad al CRAI UNEMI: %s' % reg1, request, "add")
                return JsonResponse({"result": "ok", "mensaje": u'Se registro correctamente...'})

            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al enviar los datos."})

        if action == 'addinscripcionbibliotecaexterno':
            try:
                persona = Persona.objects.get(pk=int(request.POST['id']))
                reg = RegistrarIngresoCrai(persona=persona,
                                           tiposerviciocrai_id=1,
                                           actividad='BIBLIOTECA',
                                           fecha=datetime.now().date(),
                                           horainicio=datetime.now().time())
                reg.save(request)
                log(u'Adicionó una nueva visita de incripcion crai: %s' % reg.persona, request, "add")
                return JsonResponse({"result": "ok", "mensaje": u'Se registro correctamente...'})

            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al enviar los datos."})

        if action == 'existe_inscripcion_activa':
            try:
                return JsonResponse({"result": "ok", "existe": False})
                # if 'id' in request.POST:
                #     inscripcion = Inscripcion.objects.get(pk=int(request.POST['id']))
                #     if RegistrarIngresoCrai.objects.filter(status=True, inscripcion=inscripcion, fecha=datetime.now().date(), horafin__isnull=True).exists():
                #         return JsonResponse({"result": "ok", "existe": True, "mensaje": inscripcion.persona.nombre_completo_inverso()+', ya se encuentra registrado el dia de hoy...'})
                #     else:
                #         return JsonResponse({"result": "ok", "existe": False})
                # else:
                #     return JsonResponse({"result": "bad", "mensaje": u"Error al validar los datos."})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        if action == 'existe_administrativo_activo':
                try:
                    if 'id' in request.POST:
                        distributivo = DistributivoPersona.objects.get(pk=int(request.POST['id']))
                        if RegistrarIngresoCrai.objects.filter(status=True, persona=distributivo.persona, fecha=datetime.now().date(), horafin__isnull=True).exists():
                            return JsonResponse({"result": "ok", "existe": True, "mensaje": distributivo.persona.nombre_completo_inverso() + ', ya se encuentra registrado el dia de hoy...'})
                        else:
                            return JsonResponse({"result": "ok", "existe": False})
                    else:
                        return JsonResponse({"result": "bad", "mensaje": u"Error al validar los datos."})
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})


        if action == 'existe_inscripcion_activa_externo':
            try:
                return JsonResponse({"result": "ok", "existe": False})
                # if 'id' in request.POST:
                #     persona = Persona.objects.get(pk=int(request.POST['id']))
                #     if RegistrarIngresoCrai.objects.filter(status=True, persona=persona, fecha=datetime.now().date(), horafin__isnull=True).exists():
                #         return JsonResponse({"result": "ok", "existe": True, "mensaje": persona.nombre_completo_inverso()+', ya se encuentra registrado el dia de hoy...'})
                #     else:
                #         return JsonResponse({"result": "ok", "existe": False})
                # else:
                #     return JsonResponse({"result": "bad", "mensaje": u"Error al validar los datos."})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        if action == 'delvisita':
            try:
                id=int(encrypt(request.POST['id']))
                visita = RegistrarIngresoCrai.objects.get(pk=id)
                visita.status=False
                visita.save(request)
                log(u'Elimino el registro ingreso CRAI: %s' % visita.persona if visita.persona else visita.inscripcion.persona, request, "add")
                res_json = {"error": False, "mensaje": 'Registro eliminado'}
            except Exception as ex:
                res_json = {'error': True, "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)

        if action == 'reportegeneral_pdf':
            try:
                if 'de' in request.POST and 'hasta' in request.POST:
                    data['fecha'] = datetime.now().date()
                    data['fechade'] = convertir_fecha(request.POST['de'])
                    data['fechahasta'] = convertir_fecha(request.POST['hasta'])
                    data['cantidad_estudiantes'] = cant_est =  RegistrarIngresoCrai.objects.filter(persona__isnull=True, fecha__range=(convertir_fecha(request.POST['de']), convertir_fecha(request.POST['hasta'])), status= True).count()
                    data['cantidad_administrativo'] =cant_adm =  RegistrarIngresoCrai.objects.filter(inscripcion__isnull=True, regimenlaboral__id=1, fecha__range=(convertir_fecha(request.POST['de']), convertir_fecha(request.POST['hasta'])), status= True).count()
                    data['cantidad_docententes'] = cant_doc =  RegistrarIngresoCrai.objects.filter(inscripcion__isnull=True, regimenlaboral__id=2, fecha__range=(convertir_fecha(request.POST['de']), convertir_fecha(request.POST['hasta'])), status=True).count()
                    data['cantidad_trabajadores'] = cant_tra =  RegistrarIngresoCrai.objects.filter(inscripcion__isnull=True, regimenlaboral__id=3, fecha__range=(convertir_fecha(request.POST['de']), convertir_fecha(request.POST['hasta'])), status=True).count()
                    data['total'] = cant_est + cant_adm + cant_doc + cant_tra
                    return conviert_html_to_pdf(
                        'adm_gimnasio/reportegeneral_pdf.html',
                        {
                            'pagesize': 'A4',
                            'data': data,
                        }
                    )
            except Exception as ex:
                pass

        if action == 'reportevisita_history_pdf':
            try:
                if 'de' in request.POST and 'hasta' in request.POST:
                    data['fechade'] = convertir_fecha(request.POST['de'])
                    data['fechahasta'] = convertir_fecha(request.POST['hasta'])
                    if 'tipo' in request.POST and int(request.POST['tipo'])>0:
                        if int(request.POST['tipo']) == 1:
                            data['administrativos'] = administradores = RegistrarIngresoCrai.objects.filter(inscripcion__isnull=True, regimenlaboral__id=1, fecha__range=(convertir_fecha(request.POST['de']), convertir_fecha(request.POST['hasta'])), status=True)
                        elif int(request.POST['tipo']) == 2:
                            data['trabajadores'] = trabajadores = RegistrarIngresoCrai.objects.filter(inscripcion__isnull=True, regimenlaboral__id=4, fecha__range=(convertir_fecha(request.POST['de']), convertir_fecha(request.POST['hasta'])), status=True)
                        elif int(request.POST['tipo']) == 3:
                            data['docentes'] = docentes = RegistrarIngresoCrai.objects.filter(inscripcion__isnull=True, regimenlaboral__id=2, fecha__range=(convertir_fecha(request.POST['de']), convertir_fecha(request.POST['hasta'])),status=True)
                        elif int(request.POST['tipo']) == 5:
                            data['externos'] = externos = RegistrarIngresoCrai.objects.filter(inscripcion__isnull=True, regimenlaboral_id__isnull=True, fecha__range=(convertir_fecha(request.POST['de']), convertir_fecha(request.POST['hasta'])),status=True)
                        else:
                            data['estudiantes'] = estudiantes = RegistrarIngresoCrai.objects.filter(inscripcion_id__isnull=False, regimenlaboral_id__isnull=True, fecha__range=(convertir_fecha(request.POST['de']), convertir_fecha(request.POST['hasta'])),status=True)
                    else:
                        data['fecha'] = datetime.now().date()
                        data['estudiantes'] = estudiantes =  RegistrarIngresoCrai.objects.filter(inscripcion_id__isnull=False, regimenlaboral_id__isnull=True,fecha__range=(convertir_fecha(request.POST['de']), convertir_fecha(request.POST['hasta'])), status= True)
                        data['administrativos'] = administradores =  RegistrarIngresoCrai.objects.filter(inscripcion__isnull=True, regimenlaboral__id=1, fecha__range=(convertir_fecha(request.POST['de']), convertir_fecha(request.POST['hasta'])), status= True)
                        data['docentes'] = docentes =  RegistrarIngresoCrai.objects.filter(inscripcion__isnull=True, regimenlaboral__id=2, fecha__range=(convertir_fecha(request.POST['de']), convertir_fecha(request.POST['hasta'])), status=True)
                        data['trabajadores'] = trabajadores =  RegistrarIngresoCrai.objects.filter(inscripcion__isnull=True, regimenlaboral__id=4, fecha__range=(convertir_fecha(request.POST['de']), convertir_fecha(request.POST['hasta'])), status=True)
                        data['externos'] = externos =  RegistrarIngresoCrai.objects.filter(inscripcion__isnull=True, regimenlaboral_id__isnull=True, fecha__range=(convertir_fecha(request.POST['de']), convertir_fecha(request.POST['hasta'])), status=True)



                data['tipo'] = int(request.POST['tipo'])
                data['fecha'] = datetime.now().date()
                return conviert_html_to_pdf(
                    'adm_crai/reporte_historypdf.html',
                    {
                        'pagesize': 'A4 landscape',
                        'data': data,
                    }
                )
            except Exception as ex:
                pass

        if action == 'reportevisita_history2_pdf':
            try:
                if 'de' in request.POST and 'hasta' in request.POST:
                    data['fechade'] = convertir_fecha(request.POST['de'])
                    data['fechahasta'] = convertir_fecha(request.POST['hasta'])
                    if 'tipo' in request.POST and int(request.POST['tipo'])>0:
                        data['fecha'] = datetime.now().date()
                        data['estudiantes'] = estudiantes =  RegistrarIngresoCrai.objects.filter(inscripcion_id__isnull=False,regimenlaboral_id__isnull=True, tiposerviciocrai=int(request.POST['tipo']),fecha__range=(convertir_fecha(request.POST['de']), convertir_fecha(request.POST['hasta'])), status= True)
                        data['administrativos'] = administradores =  RegistrarIngresoCrai.objects.filter(inscripcion__isnull=True, regimenlaboral__id=1, tiposerviciocrai=int(request.POST['tipo']),fecha__range=(convertir_fecha(request.POST['de']), convertir_fecha(request.POST['hasta'])), status= True)
                        data['docentes'] = docentes =  RegistrarIngresoCrai.objects.filter(inscripcion__isnull=True, regimenlaboral__id=2, tiposerviciocrai=int(request.POST['tipo']),fecha__range=(convertir_fecha(request.POST['de']), convertir_fecha(request.POST['hasta'])), status=True)
                        data['trabajadores'] = trabajadores =  RegistrarIngresoCrai.objects.filter(inscripcion__isnull=True, regimenlaboral__id=3, tiposerviciocrai=int(request.POST['tipo']),fecha__range=(convertir_fecha(request.POST['de']), convertir_fecha(request.POST['hasta'])), status=True)
                        data['externos'] = externos =  RegistrarIngresoCrai.objects.filter(inscripcion__isnull=True, regimenlaboral_id__isnull=True, tiposerviciocrai=int(request.POST['tipo']),fecha__range=(convertir_fecha(request.POST['de']), convertir_fecha(request.POST['hasta'])), status=True)



                data['tipo'] = int(request.POST['tipo'])
                data['fecha'] = datetime.now().date()
                return conviert_html_to_pdf(
                    'adm_crai/reporte_history_2_pdf.html',
                    {
                        'pagesize': 'A4 landscape',
                        'data': data,
                    }
                )
            except Exception as ex:
                pass

        if action == 'reqistrar_salida':
            try:
                if not 'id' in request.POST:
                    raise NameError('Error al registrar salida')
                visita = RegistrarIngresoCrai.objects.get(pk=int(encrypt(request.POST['id'])))
                visita.horafin = datetime.now().time()
                visita.save(request)
                messages.success(request, f'Salida marcada con existo: {visita.horafin.strftime("%I:%M %p")}')
                log(u'Registro hora de salida del CRAI: %s' % visita.persona if visita.persona else visita.inscripcion.persona, request, "add")
                return JsonResponse({"result": True, "horafin": visita.horafin.strftime("%H:%M ")})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": False, "mensaje": u"Error al guardar los datos."})

        if action == 'reqistrar_observacion':
            try:
                if 'id' in request.POST and 'observacion' in request.POST:
                    visita = RegistrarIngresoCrai.objects.get(pk=int(request.POST['id']))
                    visita.observacion = request.POST['observacion']
                    # tiene_horafin = False
                    # if not visita.horafin:
                    #     visita.horafin = request.POST['horafin']
                    #     tiene_horafin = True
                    visita.save(request)
                    log(u'Registro observacion a la visita de la persona: %s' % visita.persona if visita.persona else visita.inscripcion.persona, request, "add")
                    # return JsonResponse({"result": "ok", "observacion": visita.observacion, "horafin": str(visita.horafin), "tiene_horafin": tiene_horafin })
                    return JsonResponse({"result": "ok", "observacion": visita.observacion})
                else:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al registrar la observación."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        if action == 'addexterno':
            try:
                f = ClienteExternoCraiForm(request.POST)
                if f.is_valid():
                    if f.cleaned_data['cedula'] and Persona.objects.filter(cedula=f.cleaned_data['cedula']).exists():
                        return JsonResponse({"result": "bad", "mensaje": u"El numero de cedula ya esta registrado."})
                    clienteexterno = Persona(nombres=f.cleaned_data['nombres'],
                                             apellido1=f.cleaned_data['apellido1'],
                                             apellido2=f.cleaned_data['apellido2'],
                                             cedula=f.cleaned_data['cedula'],
                                             sexo=f.cleaned_data['sexo'],
                                             tipopersona=1,
                                             direccion=f.cleaned_data['direccion'],
                                             direccion2=f.cleaned_data['direccion2'],
                                             nacimiento=datetime.now().date())
                    clienteexterno.save(request)
                    externo = Externo(persona=clienteexterno)
                    externo.save(request)
                    clienteexterno.crear_perfil(externo=externo)
                    clienteexterno.mi_perfil()
                    log(u'Adiciono cliente externo: %s' % clienteexterno, request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                     raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        if action == 'addregistro':
            try:
                tipo=int(request.POST.get('tiporegistro', 0))
                tiposervicio=int(request.POST.get('tiposerviciocrai', 0))
                form = RegistroEstudianteForm(request.POST)
                if not tipo == 1:
                    form.fields['profesor'].required = False

                if form.is_valid():
                    if tipo == 1:
                        reg = RegistrarIngresoCrai(inscripcion=form.cleaned_data['inscripcion'],
                                                   persona=form.cleaned_data['persona'],
                                                   profesor=form.cleaned_data['profesor'],
                                                   tiposerviciocrai= form.cleaned_data['tiposerviciocrai'],
                                                   actividad= form.cleaned_data['actividad'],
                                                   cubiculo=form.cleaned_data['cubiculo'],
                                                   fecha=datetime.now().date(),
                                                   horainicio=datetime.now().time())
                    else:
                        reg = RegistrarIngresoCrai(inscripcion=form.cleaned_data['inscripcion'],
                                                   tiposerviciocrai=form.cleaned_data['tiposerviciocrai'],
                                                   librokoha=form.cleaned_data['libro'],
                                                   cubiculo=form.cleaned_data['cubiculo'],
                                                   fecha=datetime.now().date(),
                                                   horainicio=datetime.now().time())
                    reg.save(request)
                    log(u'Adicionó una nueva visita de incripcion crai: %s' % reg.inscripcion.persona, request, "add")
                    return JsonResponse({"result": False}, safe=False)
                else:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()],
                                         "mensaje": "Error en el formulario"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde.".format(str(ex))}, safe=False)

        if action == 'addregistrodocente':
            try:
                form = RegistroDocenteAdministrativoForm(request.POST)
                if form.is_valid():
                    reg = RegistrarIngresoCrai(persona=form.cleaned_data['funcionario'].persona,
                                               tiposerviciocrai=form.cleaned_data['tiposerviciocrai'],
                                               cubiculo=form.cleaned_data['cubiculo'],
                                               actividad=form.cleaned_data['actividad'],
                                               horasalida=form.cleaned_data['horasalida'],
                                               regimenlaboral=form.cleaned_data['funcionario'].regimenlaboral,
                                               fecha=datetime.now().date(),
                                               horainicio=datetime.now().time())
                    reg.save(request)
                    messages.success(request, 'Se registro correctamente')
                    log(u'Adicionó una nueva visita de: %s' % reg.persona, request, "add")
                    return JsonResponse({"result": False}, safe=False)
                else:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()],
                                         "mensaje": "Error en el formulario"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde.".format(str(ex))}, safe=False)

        if action == 'addregistroexterno':
            try:
                tipo=int(request.POST['tiporegistro'])
                form = RegistroExternoForm(request.POST)
                if tipo == 2:
                    form.fields['profesor'].required=False
                if form.is_valid():
                    if tipo == 1:
                        reg = RegistrarIngresoCrai(persona=form.cleaned_data['persona'],
                                                   profesor=form.cleaned_data['profesor'],
                                                   tiposerviciocrai_id=6,
                                                   actividad='VISITA PROFESOR',
                                                   fecha=datetime.now().date(),
                                                   horainicio=datetime.now().time())
                    elif tipo == 2:
                        reg = RegistrarIngresoCrai(persona=form.cleaned_data['persona'],
                                                   tiposerviciocrai_id=1,
                                                   actividad='BIBLIOTECA',
                                                   fecha=datetime.now().date(),
                                                   horainicio=datetime.now().time())

                    reg.save(request)
                else:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()],
                                         "mensaje": "Error en el formulario"})
                log(u'Adicionó una nueva visita de incripcion crai: %s' % reg.persona, request, "add")
                messages.success(request, 'Se registro correctamente')
                return JsonResponse({"result": False}, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde.".format(str(ex))}, safe=False)

        if action == 'addcubiculo':
            try:
                f = CubiculoForm(request.POST)
                if not f.is_valid():
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()],
                                         "mensaje": "Error en el formulario"})
                cubiculo = CubiculoCrai(numero=f.cleaned_data['numero'],
                                         # nombre=f.cleaned_data['nombre'],
                                         piso=f.cleaned_data['piso'],
                                         tiempo=f.cleaned_data['tiempo'],
                                         activo=f.cleaned_data['activo'])
                cubiculo.save(request)
                log(u'Adiciono cubículo: %s' % cubiculo, request, "add")
                return JsonResponse({"result": False})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": f"Error: {ex}"})

        if action == 'editcubiculo':
            try:
                cubiculo = CubiculoCrai.objects.get(id=encrypt_id(request.POST['id']))
                f = CubiculoForm(request.POST, instancia=cubiculo)
                if not f.is_valid():
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()],
                                         "mensaje": "Error en el formulario"})
                # cubiculo.nombre=f.cleaned_data['nombre']
                cubiculo.numero=f.cleaned_data['numero']
                cubiculo.piso=f.cleaned_data['piso']
                cubiculo.tiempo=f.cleaned_data['tiempo']
                cubiculo.activo=f.cleaned_data['activo']
                cubiculo.save(request)
                log(u'Edito cubículo: %s' % cubiculo, request, "edit")
                return JsonResponse({"result": False})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": u"Error al guardar los datos."})

        if action == 'delcubiculo':
            try:
                id = encrypt_id(request.POST['id'])
                cubiculo = CubiculoCrai.objects.get(pk=id)
                cubiculo.status = False
                cubiculo.save(request)
                log(u'Elimino cubículo : %s' % cubiculo, request, "del")
                res_json = {"error": False, "mensaje": 'Registro eliminado'}
            except Exception as ex:
                res_json = {'error': True, "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)

        elif action == 'cambiarctivo':
            with transaction.atomic():
                try:
                    registro = CubiculoCrai.objects.get(pk=encrypt_id(request.POST['id']))
                    registro.activo = eval(request.POST['val'].capitalize())
                    registro.save(request)
                    log(u'Cambio estado activo de cubiculo: %s (%s)' % (registro, registro.activo), request, "edit")
                    return JsonResponse({"result": True, 'mensaje': 'Cambios guardados'})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": False})
        return JsonResponse({"result": "bad", "mensaje": u"Solicitud Incorrecta."})
    else:
        if 'action' in request.GET:
            data['action'] = action = request.GET['action']

            if action == 'addinscripcion':  #----sin utilizar
                try:
                    data['title'] = u'Registrar Ingreso CRAI'
                    data['s'] = request.GET['s']
                    data['id'] = request.GET['id']
                    # formulario = RegistrarIngresoCraiForm()
                    data['form'] = RegistrarIngresoCraiForm()
                    persona = Persona.objects.get(inscripcion=int(request.GET['id']))
                    data['nombre'] = persona.nombre_completo_inverso()
                    return render(request, "adm_crai/addinscripcion.html", data)
                    # template = get_template("adm_crai/addinscripcion.html")
                    # json_content = template.render(data)
                    # return JsonResponse({"result": "ok", 'formulario': RegistrarIngresoCraiForm(),'nombre': persona.nombre_completo_inverso()}, content_type="application/json")
                except Exception as ex:
                    pass
                    # transaction.set_rollback(True)
                    # return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

            if action == 'addinscripcionbiblioteca':
                try:
                    data['title'] = u'Registrar Ingreso CRAI Biblioteca'
                    data['inscripcion'] = Inscripcion.objects.get(pk=int(request.GET['idi']))
                    return render(request, "adm_crai/addinscripcionbiblioteca.html", data)
                except Exception as ex:
                    pass

            if action == 'addinscripcionbibliotecaexterno':
                try:
                    data['title'] = u'Registrar Ingreso CRAI Biblioteca'
                    data['persona'] = Persona.objects.get(pk=int(request.GET['idi']))
                    return render(request, "adm_crai/addinscripcionbibliotecaexterno.html", data)
                except Exception as ex:
                    pass

            if action == 'addinscripcionexterno':
                try:
                    data['title'] = u'Registrar Ingreso CRAI'
                    data['s'] = request.GET['s']
                    data['id'] = request.GET['id']
                    form = RegistrarIngresoCraiForm()
                    data['form'] = form
                    return render(request, "adm_crai/addinscripcionexterno.html", data)
                except Exception as ex:
                    pass

            if action == 'delvisita':
                try:
                    data['title'] = u'Eliminar Registro Ingreso'
                    data['visita'] =  RegistrarIngresoCrai.objects.get(pk=int(request.GET['id']))
                    return render(request, "adm_crai/delvisita.html", data)
                except Exception as ex:
                    pass

            elif action == 'registrarinscripcion':
                data['title'] = u'Control de acceso al CRAI UNEMI'
                data['title1'] = u'Listado de estudiantes UNEMI'
                try:
                    search = None
                    ids = None
                    inscripciones = []
                    if 's' in request.GET:
                        search = request.GET['s'].strip()
                        ss = search.split(' ')
                        if len(ss) == 1:
                            inscripciones = Inscripcion.objects.filter(Q(persona__nombres__icontains=search) |
                                                                       Q(persona__apellido1__icontains=search) |
                                                                       Q(persona__apellido2__icontains=search) |
                                                                       Q(persona__cedula__icontains=search) |
                                                                       Q(persona__pasaporte__icontains=search) |
                                                                       Q(identificador__icontains=search) |
                                                                       Q(inscripciongrupo__grupo__nombre__icontains=search) |
                                                                       Q(persona__usuario__username__icontains=search))
                            # .exclude(registraringresocrai__fecha=datetime.now().date(), registraringresocrai__horafin__isnull=True)
                        else:
                            inscripciones = Inscripcion.objects.filter((Q(persona__nombres__icontains=ss[0]) & Q(persona__nombres__icontains=ss[1])) | (
                                Q(persona__apellido1__icontains=ss[0]) & Q(persona__apellido2__icontains=ss[1])))

                    paging = MiPaginador(inscripciones, 15)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['inscripciones'] = page.object_list
                    data['tiposervicio'] = TipoServicioCrai.objects.filter(status=True)
                    # data['profesor'] = Profesor.objects.filter(status=True, activo=True)
                    # data['libros'] = LibroKohaProgramaAnaliticoAsignatura.objects.filter(status=True)
                    data['actividadescrais'] = ActividadesCrai.objects.filter(status=True).order_by('orden')
                    return render(request, "adm_crai/view_registrarinscripcion.html", data)
                except Exception as ex:
                    pass

            if action == 'reportevisita_estudiante_excel':
                try:
                    fechadesde = request.GET['de']
                    fechahasta = request.GET['hasta']

                    __author__ = 'Unemi'
                    styrowD = easyxf('font: name Times New Roman, color-index black, bold off; borders: left thin, right thin, top thin, bottom thin')
                    styrow = easyxf('font: name Times New Roman, color-index black, bold off; alignment: horiz centre; borders: left thin, right thin, top thin, bottom thin')
                    style_col = easyxf('font: name Times New Roman, color-index black, bold on; alignment: horiz centre; borders: left thin, right thin, top thin, bottom thin')
                    style_sb1 = easyxf('font: name Times New Roman, color-index black, bold on')
                    style_sb = easyxf('font: name Times New Roman, color-index black, bold off')
                    title = easyxf( 'font: name Times New Roman, color-index green, bold on , height 350; alignment: horiz centre')
                    title1 = easyxf( 'font: name Times New Roman, color-index green, bold on , height 250; alignment: horiz centre')
                    style_date = easyxf('borders: left thin, right thin, top thin, bottom thin; alignment: horiz centre', num_format_str='yy/mm/dd')
                    style_hr = easyxf('borders: left thin, right thin, top thin, bottom thin; alignment: horiz centre', num_format_str='h:mm')

                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('Reporte Visita')

                    ws.write_merge(0, 0, 0, 9, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                    ws.write_merge(1, 1, 0, 9, 'REPORTE DE ESTUDIANTES ATENDIDOS', title1)
                    ws.write_merge(3, 3, 1, 1, 'PERIODO:  ', style_sb1)
                    ws.write_merge(3, 3, 2, 2, 'DESDE  ' + request.GET['de'] + ' HASTA  ' + request.GET['hasta'], style_sb)

                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=ReporteVisitaEstudiante ' + random.randint(1, 10000).__str__() + '.xls'

                    columns = [
                        (u"N°", 1100),
                        (u"N° CÉDULA", 4500),
                        (u"APELLIDOS Y NOMBRES", 10000),
                        (u"FACULTAD", 10000),
                        (u"CARRERA", 10000),
                        (u"FECHA", 4500),
                        (u"HORA INGRESO", 4500),
                        (u"HORA SALIDA", 4500),
                        (u"SERVICIO", 5000),
                        (u"ACTIVIDAD", 7000),
                    ]
                    row_num = 5
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], style_col)
                        ws.col(col_num).width = columns[col_num][1]
                    # persona__estudioinscripcion__isnull=False
                    visitas = RegistrarIngresoCrai.objects.filter(status=True,inscripcion_id__isnull=False, regimenlaboral_id__isnull=True,
                                                                  fecha__range=(convertir_fecha(fechadesde), convertir_fecha(fechahasta)) ).order_by('persona__apellido1')

                    row_num = 6
                    cont = 1

                    for vis in visitas:

                        ws.write(row_num, 0, cont, styrow)
                        ws.write(row_num, 1, vis.inscripcion.persona.cedula, styrowD)
                        ws.write(row_num, 2, str(vis.inscripcion.persona.nombre_completo_inverso()) , styrowD)
                        ws.write(row_num, 3, str(vis.inscripcion.coordinacion.nombre), styrow)
                        ws.write(row_num, 4, str(vis.inscripcion.carrera.nombre), styrow)
                        ws.write(row_num, 5, vis.fecha, style_date)
                        ws.write(row_num, 6, vis.horainicio, style_hr)
                        ws.write(row_num, 7, vis.horafin, style_hr)
                        ws.write(row_num, 8, str(vis.tiposerviciocrai.descripcion), styrow)
                        ws.write(row_num, 9, str(vis.actividad), styrow)
                        row_num += 1
                        cont += 1

                    row_num += 1
                    ws.write(row_num, 8, 'TOTAL', styrow)
                    ws.write(row_num, 9, visitas.count(), styrow)
                    wb.save(response)
                    return response
                except Exception as ex:
                    pass

            if action == 'reportevisita_estudiante_pdf':
                try:
                    fechadesde = request.GET['de']
                    fechahasta = request.GET['hasta']
                    visitas = RegistrarIngresoCrai.objects.filter(status=True, inscripcion_id__isnull=False,regimenlaboral_id__isnull=True,
                                                                  fecha__range=(convertir_fecha(fechadesde),
                                                                                convertir_fecha(fechahasta)) ).order_by('persona__apellido1')
                    data['visitasEst'] = visitas

                    data['total'] = visitas.count()
                    data['desde'] = fechadesde
                    data['hasta'] = fechahasta
                    data['fechahoy'] = datetime.now().date()
                    return conviert_html_to_pdf(
                        'adm_crai/reporte_estudiantes_pdf.html',
                        {
                            'pagesize': 'A4 landscape',
                            'data': data,
                        }
                    )
                except Exception as ex:
                    pass


            if action == 'verhistorial':
                data['title'] = u'Control de acceso al CRAI UNEMI (Historial)'
                try:
                    search = None
                    ids = None
                    fecha = None
                    tipov = None
                    visitas = []
                    visitas = RegistrarIngresoCrai.objects.filter(status=True, fecha=datetime.now().date()).order_by('-horainicio')
                    if 's' in request.GET:
                        search = request.GET['s'].strip()
                        ss = search.split(' ')
                        if 'fecha' in request.GET:
                            fecha = convertir_fecha(request.GET['fecha'])
                            if len(ss) == 1:
                                visitas = RegistrarIngresoCrai.objects.filter(Q(status=True),
                                                                              ((Q(persona__nombres__icontains=search) |
                                                                                Q(persona__apellido1__icontains=search) |
                                                                                Q(persona__apellido2__icontains=search) |
                                                                                Q(persona__cedula__icontains=search)) | (
                                                                                           Q(inscripcion__persona__nombres__icontains=search) |
                                                                                           Q(inscripcion__persona__apellido1__icontains=search) |
                                                                                           Q(inscripcion__persona__apellido2__icontains=search) |
                                                                                           Q(inscripcion__persona__cedula__icontains=search))),
                                                                              Q(fecha=convertir_fecha(request.GET['fecha']))).exclude(horafin__isnull=True).order_by('horainicio')
                            else:
                                visitas = RegistrarIngresoCrai.objects.filter(Q(status=True),
                                                                              (((Q(persona__nombres__icontains=ss[0]) & Q(persona__nombres__icontains=ss[0])) |
                                                                                (Q(persona__apellido1__icontains=ss[0]) & Q(persona__apellido2__icontains=ss[1]))) | ((Q(inscripcion__persona__nombres__icontains=ss[0]) & Q(inscripcion__persona__nombres__icontains=ss[1])) |
                                                                                       (Q(inscripcion__persona__apellido1__icontains=ss[0]) & Q(inscripcion__persona__apellido2__icontains=ss[1])))), Q(fecha=convertir_fecha(request.GET['fecha']))).exclude(horafin__isnull=True).order_by('horainicio')
                        else:
                            if len(ss) == 1:
                                visitas = RegistrarIngresoCrai.objects.filter(Q(status=True),
                                                                              ((Q(persona__nombres__icontains=search) |
                                                                                Q(persona__apellido1__icontains=search) |
                                                                                Q(persona__apellido2__icontains=search) |
                                                                                Q(persona__cedula__icontains=search)) | (
                                                                                           Q(inscripcion__persona__nombres__icontains=search) |
                                                                                           Q(inscripcion__persona__apellido1__icontains=search) |
                                                                                           Q(inscripcion__persona__apellido2__icontains=search) |
                                                                                           Q(inscripcion__persona__cedula__icontains=search)))).exclude(horafin__isnull=True).order_by('horainicio')
                            else:
                                visitas = RegistrarIngresoCrai.objects.filter(Q(status=True),
                                                                              (((Q(persona__nombres__icontains=ss[0]) & Q(persona__nombres__icontains=ss[0])) |
                                                                                (Q(persona__apellido1__icontains=ss[0]) & Q(
                                                                                    persona__apellido2__icontains=ss[1]))) | (
                                                                                       (Q(inscripcion__persona__nombres__icontains=ss[0]) & Q(inscripcion__persona__nombres__icontains=ss[1])) |
                                                                                       (Q(inscripcion__persona__apellido1__icontains=ss[0]) & Q(inscripcion__persona__apellido2__icontains=ss[1]))))).exclude(horafin__isnull=True).order_by('horainicio')
                    elif 'fecha' in request.GET:
                        visitas = RegistrarIngresoCrai.objects.filter(status=True, fecha=convertir_fecha(request.GET['fecha'])).exclude(horafin__isnull=True).order_by('horainicio')
                        fecha = convertir_fecha(request.GET['fecha'])
                    # else:
                    #     visitas = RegistrarIngresoCrai.objects.filter(status=True).exclude(horafin__isnull=False).order_by('-fecha', '-horainicio')
                    paging = MiPaginador(visitas, 10)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['searchHistory'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['visitasHistorial'] = page.object_list
                    data['fechaselectHistory'] = fecha
                    data['tiposervicioHis'] = TipoServicioCrai.objects.filter(status=True)
                    data['horaHistory'] = datetime.now().time().strftime("%H:%M")
                    # return render(request, "adm_crai/verhistorial.html", data)
                    return render(request, "adm_crai/view1.html", data)
                except Exception as ex:
                    pass

            if action == 'registrarexternos':
                try:
                    data['title1'] = u'Listado de personal externo UNEMI'
                    data['title'] = u'Control de acceso al CRAI UNEMI'
                    search = None
                    ids = None
                    fecha = None
                    # unemiper = RegistrarIngresoCrai.objects.values_list('persona_id', flat=True).filter(status=True, persona__isnull=False, fecha=datetime.now().date(), horafin__isnull=True)
                    # listadist = []
                    # for gym in unemiper:
                    #     listadist.append(DistributivoPersona.objects.filter(status= True, persona=gym.persona, regimenlaboral=gym.regimenlaboral, estadopuesto__id=PUESTO_ACTIVO_ID)[0].id)
                    externos = []
                    externos = Externo.objects.filter(status=True, fecha_creacion=datetime.now().date())
                    if 's' in request.GET:
                        search = request.GET['s'].strip()
                        ss = search.split(' ')
                        if len(ss) == 1:
                            externos = Externo.objects.filter((Q(persona__nombres__icontains=search) |
                                                               Q(persona__apellido1__icontains=search) |
                                                               Q(persona__apellido2__icontains=search) |
                                                               Q(persona__cedula__icontains=search) |
                                                               Q(persona__pasaporte__icontains=search)))
                            # .exclude(persona__id__in=unemiper)
                        else:
                            externos = Externo.objects.filter((Q(persona__nombres__icontains=ss[0]) & Q(persona__nombres__icontains=ss[1]) |
                                                               Q(persona__apellido1__icontains=ss[0]) & Q(persona__apellido2__icontains=ss[1])))
                    # elif 'id' in request.GET:
                    #     externos = DistributivoPersona.objects.filter(Q(pk=int(request.GET['id']))& Q(estadopuesto__id=PUESTO_ACTIVO_ID) & (Q(regimenlaboral__id=1) | Q(regimenlaboral__id=2)|Q(regimenlaboral__id=4)))
                    # else:
                    #     externos = DistributivoPersona.objects.filter(Q(estadopuesto__id=PUESTO_ACTIVO_ID) & (Q(regimenlaboral__id=1)|Q(regimenlaboral__id=2)|Q(regimenlaboral__id=4))).exclude(id__in=listadist)
                    if 'fecha' in request.GET:
                        # externos = Externo.objects.filter(status=True, fecha_creacion__range=(convertir_fecha(request.GET['fecha']), convertir_fecha(request.GET['fecha'])))
                        externos = Externo.objects.filter(status=True, fecha_creacion__year=convertir_fecha(request.GET['fecha']).year,
                                                                            fecha_creacion__month=convertir_fecha(request.GET['fecha']).month,
                                                                        fecha_creacion__day=convertir_fecha(request.GET['fecha']).day )
                        fecha = convertir_fecha(request.GET['fecha'])


                    paging = MiPaginador(externos, 15)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['fechaselect'] = fecha #if fecha else datetime.now().date()
                    data['externos'] = page.object_list
                    return render(request, "adm_crai/view_registrarexternos.html", data)
                except Exception as ex:
                    pass

            if action == 'addexterno':
                try:
                    # puede_realizar_accion(request, 'sga.puede_modificar_administrativos')
                    data['title'] = u'Adicionar Clientes Externos'
                    form = ClienteExternoCraiForm()
                    data['form'] = form
                    return render(request, "adm_crai/addexterno.html", data)
                except Exception as ex:
                    pass

            if action == 'addaux':
                try:
                    data['title'] = u'Registrar Actividad CRAI Biblioteca'
                    data['inscripcion'] = Inscripcion.objects.get(pk=int(request.GET['idi']))
                    data['actividaid'] = int(request.GET['ida'])
                    data['ci'] = request.GET['ci']
                    return render(request, "adm_crai/addaux.html", data)
                except Exception as ex:
                    pass


            if action == 'registrardocentes_administrativos':
                try:
                    data['title1'] = u'Listado de personal Docente - administrativo'
                    data['title'] = u'Control de acceso al CRAI UNEMI'
                    search = None
                    ids = None
                    unemiper = RegistrarIngresoCrai.objects.filter(status=True, persona__isnull=False,
                                                                      fecha=datetime.now().date(), horafin__isnull=True)
                    listadist = []
                    for gym in unemiper:
                        listadist.append(DistributivoPersona.objects.filter(status=True, persona=gym.persona,
                                                                            regimenlaboral=gym.regimenlaboral,
                                                                            estadopuesto__id=PUESTO_ACTIVO_ID)[0].id)
                    if 's' in request.GET:
                        search = request.GET['s'].strip()
                        ss = search.split(' ')
                        if len(ss) == 1:
                            administrativos = DistributivoPersona.objects.filter(
                                Q(estadopuesto__id=PUESTO_ACTIVO_ID) & (
                                        Q(regimenlaboral__id=1) | Q(regimenlaboral__id=2) | Q(regimenlaboral__id=4)) &
                                (Q(persona__nombres__icontains=search) |
                                 Q(persona__apellido1__icontains=search) |
                                 Q(persona__apellido2__icontains=search) |
                                 Q(persona__cedula__icontains=search) |
                                 Q(persona__pasaporte__icontains=search) |
                                 Q(
                                     denominacionpuesto__descripcion__icontains=search))).exclude(
                                id__in=listadist)
                        else:
                            administrativos = DistributivoPersona.objects.filter(
                                Q(estadopuesto__id=PUESTO_ACTIVO_ID) & (
                                        Q(regimenlaboral__id=1) | Q(regimenlaboral__id=2) | Q(regimenlaboral__id=4)) &
                                (Q(persona__nombres__icontains=ss[0]) & Q(
                                    persona__nombres__icontains=ss[1]) |
                                 Q(persona__apellido1__icontains=ss[
                                     0]) & Q(persona__apellido2__icontains=
                                             ss[1]) |
                                 Q(
                                     denominacionpuesto__descripcion__icontains=
                                     ss[0]) & Q(
                                            denominacionpuesto__descripcion__icontains=
                                            ss[1]))).exclude(
                                id__in=listadist)
                    elif 'id' in request.GET:
                        administrativos = DistributivoPersona.objects.filter(
                            Q(pk=int(request.GET['id'])) & Q(estadopuesto__id=PUESTO_ACTIVO_ID) & (
                                    Q(regimenlaboral__id=1) | Q(regimenlaboral__id=2) | Q(regimenlaboral__id=4)))
                    else:
                        administrativos = DistributivoPersona.objects.filter(Q(estadopuesto__id=PUESTO_ACTIVO_ID) & (
                                Q(regimenlaboral__id=1) | Q(regimenlaboral__id=2) | Q(regimenlaboral__id=4))).exclude(
                            id__in=listadist)
                    paging = MiPaginador(administrativos, 15)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['administrativos'] = page.object_list
                    return render(request, "adm_crai/view_registrardocenteadmin.html", data)
                except Exception as ex:
                    pass
                #     template = get_template("adm_crai/addinscripcionbibliodocentes.html")
                #     json_content = template.render(data)
                #     return JsonResponse({"result": "ok", 'data': json_content})
                # except Exception as ex:
                #     transaction.set_rollback(True)
                #     return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

            if action == 'docentes_administrativos':
                try:

                    search = None
                    ids = None
                    tipov = None
                    administrativos = RegistrarIngresoCrai.objects.filter( Q(regimenlaboral__id=1) | Q( regimenlaboral__id=2) | Q(regimenlaboral__id=4),
                                                                    status=True, persona__isnull=False)


                    if 'fecha' in request.GET:
                        administrativos = administrativos.filter(fecha=convertir_fecha(request.GET['fecha']))
                        fecha = convertir_fecha(request.GET['fecha'])

                    if 's' in request.GET:
                        search = request.GET['s'].strip()
                        ss = search.split(' ')
                        if len(ss) == 1:
                            administrativos = administrativos.filter((Q(persona__nombres__icontains=search) |
                                 Q(persona__apellido1__icontains=search) |
                                 Q(persona__apellido2__icontains=search) |
                                 Q(persona__cedula__icontains=search) |
                                 Q(persona__pasaporte__icontains=search) |
                                 Q(denominacionpuesto__descripcion__icontains=search)))
                        else:
                            administrativos = administrativos.filter(
                                (Q(persona__nombres__icontains=ss[0]) & Q(persona__nombres__icontains=ss[1]) |
                                 Q(persona__apellido1__icontains=ss[0]) & Q(persona__apellido2__icontains= ss[1]) |
                                 Q(denominacionpuesto__descripcion__icontains=ss[0]) &
                                 Q(denominacionpuesto__descripcion__icontains=ss[1])))
                    elif 'tipo_servD' in request.GET:
                        if int(request.GET['tipo_servD']) != 0:
                            administrativos = administrativos.filter(tiposerviciocrai=int( request.GET['tipo_servD'])).order_by('-horainicio')
                        tipov = (request.GET['tipo_servD'])

                    paging = MiPaginador(administrativos, 15)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['tipov_D'] = tipov
                    data['administrativos'] = page.object_list
                #     return render(request, "adm_crai/docente_administrativo.html", data)
                # except Exception as ex:
                #     pass
                    template = get_template("adm_crai/docente_administrativo.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content, 'tipov_D': tipov})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

            if action == 'reportevisita_docenteadmi_excel':
                try:
                    fechadesde = request.GET['de']
                    fechahasta = request.GET['hasta']

                    __author__ = 'Unemi'
                    styrowD = easyxf('font: name Times New Roman, color-index black, bold off; borders: left thin, right thin, top thin, bottom thin')
                    styrow = easyxf('font: name Times New Roman, color-index black, bold off; alignment: horiz centre; borders: left thin, right thin, top thin, bottom thin')
                    style_col = easyxf('font: name Times New Roman, color-index black, bold on; alignment: horiz centre; borders: left thin, right thin, top thin, bottom thin')
                    style_sb1 = easyxf('font: name Times New Roman, color-index black, bold on')
                    style_sb = easyxf('font: name Times New Roman, color-index black, bold off')
                    title = easyxf( 'font: name Times New Roman, color-index green, bold on , height 350; alignment: horiz centre')
                    title1 = easyxf( 'font: name Times New Roman, color-index green, bold on , height 250; alignment: horiz centre')
                    style_date = easyxf('borders: left thin, right thin, top thin, bottom thin; alignment: horiz centre', num_format_str='yy/mm/dd')
                    style_hr = easyxf('borders: left thin, right thin, top thin, bottom thin; alignment: horiz centre', num_format_str='h:mm')

                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('Reporte Visita')

                    ws.write_merge(0, 0, 0, 8, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                    ws.write_merge(1, 1, 0, 8, 'REPORTE DE DOCENTES - ADMINISTRATIVOS ATENDIDOS', title1)
                    ws.write_merge(3, 3, 1, 1, 'PERIODO:  ', style_sb1)
                    ws.write_merge(3, 3, 2, 2, 'DESDE  ' + request.GET['de'] + ' HASTA  ' + request.GET['hasta'], style_sb)

                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=ReporteVisitaDocente ' + random.randint(1, 10000).__str__() + '.xls'

                    columns = [
                        (u"N°", 1100),
                        (u"N° CÉDULA", 4500),
                        (u"APELLIDOS Y NOMBRES", 10000),
                        (u"TIPO", 4500),
                        (u"FECHA", 4500),
                        (u"HORA INGRESO", 4500),
                        (u"HORA SALIDA", 4500),
                        (u"SERVICIO", 5000),
                        (u"ACTIVIDAD", 5000),
                    ]
                    row_num = 5
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], style_col)
                        ws.col(col_num).width = columns[col_num][1]

                    visitasDo = RegistrarIngresoCrai.objects.filter( Q(regimenlaboral__id=1) | Q(regimenlaboral__id=2) | Q(regimenlaboral__id=4),
                                                                status=True, persona__isnull=False,
                                                                fecha__range=(convertir_fecha(fechadesde),convertir_fecha(fechahasta)) ).order_by('persona__apellido1')

                    row_num = 6
                    cont = 1
                    for vis in visitasDo:
                        ws.write(row_num, 0, cont, styrow)
                        ws.write(row_num, 1, vis.persona.cedula, styrowD)
                        ws.write(row_num, 2, str(vis.persona.nombre_completo_inverso()), styrowD)
                        if vis.regimenlaboral.id == 1:
                            ws.write(row_num, 3, 'ADMINISTRATIVO', styrow)
                        elif vis.regimenlaboral.id == 2:
                            ws.write(row_num, 3, 'DOCENTE', styrow)
                        else:
                            ws.write(row_num, 3, 'TRABAJADOR', styrow)
                        ws.write(row_num, 4, vis.fecha, style_date)
                        ws.write(row_num, 5, vis.horainicio, style_hr)
                        ws.write(row_num, 6, vis.horafin, style_hr)
                        ws.write(row_num, 7, str(vis.tiposerviciocrai.descripcion), styrow)
                        ws.write(row_num, 8, str(vis.actividad), styrow)
                        row_num += 1
                        cont += 1

                    row_num += 1
                    ws.write(row_num, 7, 'TOTAL', styrow)
                    ws.write(row_num, 8, visitasDo.count(), styrow)
                    wb.save(response)
                    return response
                except Exception as ex:
                    pass

            if action == 'reportevisita_docenteadmi_pdf':
                try:
                    fechadesde = request.GET['de']
                    fechahasta = request.GET['hasta']
                    visitasDo = RegistrarIngresoCrai.objects.filter(Q(regimenlaboral__id=1) | Q(regimenlaboral__id=2) | Q(regimenlaboral__id=4),
                        status=True, persona__isnull=False,
                        fecha__range=(convertir_fecha(fechadesde), convertir_fecha(fechahasta))).order_by('persona__apellido1')
                    data['visitasDo'] = visitasDo

                    data['total'] = visitasDo.count()
                    data['desde'] = fechadesde
                    data['hasta'] = fechahasta
                    data['fechahoy'] = datetime.now().date()
                    return conviert_html_to_pdf(
                        'adm_crai/reporte_docenteadmi_pdf.html',
                        {
                            'pagesize': 'A4 landscape',
                            'data': data,
                        }
                    )
                except Exception as ex:
                    pass


            if action == 'personal_externo':
                data['title'] = u'Control de acceso al CRAI UNEMI'
                try:
                    search = None
                    ids = None
                    fecha = None
                    tipov = None
                    visitas = []
                    exter = Externo.objects.values_list('persona__id', flat=True).filter(status=True)
                    visitasExt = RegistrarIngresoCrai.objects.filter(status=True, fecha=datetime.now().date(), inscripcion__isnull=True, regimenlaboral_id__isnull=True,
                                                                     persona_id__in=exter).exclude(horafin__isnull=False).order_by('-horainicio')
                    if 's' in request.GET:
                        search = request.GET['s'].strip()
                        ss = search.split(' ')
                        if 'fecha' in request.GET:
                            fecha = convertir_fecha(request.GET['fecha'])
                            if len(ss) == 1:
                                visitas = visitasExt.filter(Q(status=True),
                                                                              ((Q(persona__nombres__icontains=search) |
                                                                                Q(persona__apellido1__icontains=search) |
                                                                                Q(persona__apellido2__icontains=search) |
                                                                                Q(persona__cedula__icontains=search)) ),
                                                                              Q(fecha=convertir_fecha(request.GET['fecha']))).order_by('horainicio')
                            #     .exclude(horafin__isnull=False)
                            else:
                                visitas = visitasExt.filter(Q(status=True),
                                                                              (((Q(persona__nombres__icontains=ss[0]) & Q(
                                                                                  persona__nombres__icontains=ss[0])) |
                                                                                (Q(persona__apellido1__icontains=ss[0]) & Q(
                                                                                    persona__apellido2__icontains=ss[1]))) ),
                                                                              Q(fecha=convertir_fecha(request.GET['fecha']))).order_by('horainicio')
                                # .exclude(horafin__isnull=False)
                        else:
                            if len(ss) == 1:
                                visitas =visitasExt.filter(Q(status=True),
                                                                              ((Q(persona__nombres__icontains=search) |
                                                                                Q(persona__apellido1__icontains=search) |
                                                                                Q(persona__apellido2__icontains=search) |
                                                                                Q(persona__cedula__icontains=search)) )).order_by('horainicio')
                                # .exclude(horafin__isnull=False)
                            else:
                                visitas = visitasExt.filter(Q(status=True),
                                                                              (((Q(persona__nombres__icontains=ss[0]) & Q(
                                                                                  persona__nombres__icontains=ss[0])) |
                                                                                (Q(persona__apellido1__icontains=ss[0]) & Q(
                                                                                    persona__apellido2__icontains=ss[1]))) )).order_by('horainicio')
                                # .exclude(horafin__isnull=False)
                    elif 'fecha' in request.GET:
                        visitas = RegistrarIngresoCrai.objects.filter(status=True, inscripcion_id__isnull=True, regimenlaboral_id__isnull=True,
                                            horafin__isnull=True,fecha=convertir_fecha(request.GET['fecha'])).exclude(horafin__isnull=False).order_by('-horainicio')
                        fecha = convertir_fecha(request.GET['fecha'])
                    elif 'tipo_serv_Ex' in request.GET:
                        if int(request.GET['tipo_serv_Ex']) != 0:
                            visitas = RegistrarIngresoCrai.objects.filter(status=True, tiposerviciocrai=int( request.GET['tipo_serv_Ex']),
                                          inscripcion_id__isnull=True,regimenlaboral_id__isnull=False).exclude(
                                          horafin__isnull=False).exclude(horafin__isnull=False).order_by('-horainicio')
                        tipov = int(request.GET['tipo_serv_Ex'])
                    else:
                        visitas = RegistrarIngresoCrai.objects.filter(status=True, inscripcion__isnull=True, regimenlaboral_id__isnull=True,
                                                                      fecha=datetime.now().date(), persona_id__in=exter).exclude(horafin__isnull=False).order_by('-horainicio')
                        # .exclude(horafin__isnull=False)
                    paging = MiPaginador(visitas, 10)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['searchExt'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['tiposerv_Ex'] = tipov
                    data['visitasExternos'] = page.object_list
                    data['fechaselectExt'] = fecha  # if fecha else datetime.now().date()
                    data['tiposervicioHis'] = TipoServicioCrai.objects.filter(status=True)
                    data['horaExt'] = datetime.now().time().strftime("%H:%M")
                    return render(request, "adm_crai/view1.html", data)
                except Exception as ex:
                    pass

            if action == 'reportevisita_externos_excel':
                try:
                    fechadesde = request.GET['de']
                    fechahasta = request.GET['hasta']

                    __author__ = 'Unemi'
                    styrowD = easyxf('font: name Times New Roman, color-index black, bold off; borders: left thin, right thin, top thin, bottom thin')
                    styrow = easyxf('font: name Times New Roman, color-index black, bold off; alignment: horiz centre; borders: left thin, right thin, top thin, bottom thin')
                    style_col = easyxf('font: name Times New Roman, color-index black, bold on; alignment: horiz centre; borders: left thin, right thin, top thin, bottom thin')
                    style_sb1 = easyxf('font: name Times New Roman, color-index black, bold on')
                    style_sb = easyxf('font: name Times New Roman, color-index black, bold off')
                    title = easyxf( 'font: name Times New Roman, color-index green, bold on , height 350; alignment: horiz centre')
                    title1 = easyxf( 'font: name Times New Roman, color-index green, bold on , height 250; alignment: horiz centre')
                    style_date = easyxf('borders: left thin, right thin, top thin, bottom thin; alignment: horiz centre', num_format_str='yy/mm/dd')
                    style_hr = easyxf('borders: left thin, right thin, top thin, bottom thin; alignment: horiz centre', num_format_str='h:mm')

                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('Reporte Visita')

                    ws.write_merge(0, 0, 0, 7, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                    ws.write_merge(1, 1, 0, 7, 'REPORTE DE ATENCIONES A USUARIOS EXTERNOS', title1)
                    ws.write_merge(3, 3, 1, 1, 'PERIODO:  ', style_sb1)
                    ws.write_merge(3, 3, 2, 2, 'DESDE  ' + request.GET['de'] + ' HASTA  ' + request.GET['hasta'], style_sb)

                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=ReporteVisitaExternos ' + random.randint(1, 10000).__str__() + '.xls'

                    columns = [
                        (u"N°", 1100),
                        (u"N° CÉDULA", 4500),
                        (u"APELLIDOS Y NOMBRES", 10000),
                        (u"FECHA", 4500),
                        (u"HORA INGRESO", 4500),
                        (u"HORA SALIDA", 4500),
                        (u"SERVICIO", 5000),
                        (u"ACTIVIDAD", 5000),
                    ]
                    row_num = 5
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], style_col)
                        ws.col(col_num).width = columns[col_num][1]

                    visitasEx = RegistrarIngresoCrai.objects.filter(inscripcion_id__isnull=True,regimenlaboral_id__isnull=True,
                        status=True, persona__isnull=False,
                        fecha__range=(convertir_fecha(fechadesde), convertir_fecha(fechahasta))).order_by('persona__apellido1')

                    row_num = 6
                    cont = 1

                    for vis in visitasEx:
                        ws.write(row_num, 0, cont, styrow)
                        ws.write(row_num, 1, vis.persona.cedula, styrowD)
                        ws.write(row_num, 2, str(vis.persona.nombre_completo_inverso()), styrowD)
                        ws.write(row_num, 3, vis.fecha, style_date)
                        ws.write(row_num, 4, vis.horainicio, style_hr)
                        ws.write(row_num, 5, vis.horafin, style_hr)
                        ws.write(row_num, 6, str(vis.tiposerviciocrai.descripcion), styrow)
                        ws.write(row_num, 7, str(vis.actividad), styrow)
                        row_num += 1
                        cont += 1

                    row_num += 1
                    ws.write(row_num, 6, 'TOTAL', styrow)
                    ws.write(row_num, 7, visitasEx.count(), styrow)
                    wb.save(response)
                    return response
                except Exception as ex:
                    pass

            if action == 'reportevisita_externos_pdf':
                try:
                    fechadesde = request.GET['de']
                    fechahasta = request.GET['hasta']
                    visitasEx = RegistrarIngresoCrai.objects.filter(inscripcion_id__isnull=True,
                                                regimenlaboral_id__isnull=True,
                                                status=True, persona__isnull=False,
                                                fecha__range=(convertir_fecha(fechadesde), convertir_fecha(fechahasta))).order_by('persona__apellido1')
                    data['visitasEx'] = visitasEx

                    data['total'] = visitasEx.count()
                    data['desde'] = fechadesde
                    data['hasta'] = fechahasta
                    data['fechahoy'] = datetime.now().date()
                    return conviert_html_to_pdf(
                        'adm_crai/reporte_externos_pdf.html',
                        {
                            'pagesize': 'A4 landscape',
                            'data': data,
                        }
                    )
                except Exception as ex:
                    pass

            #Nuevas acciones
            elif action == 'addregistro':
                try:
                    form=RegistroEstudianteForm()
                    form.cargar()
                    data['form']=form
                    template = get_template("adm_crai/modal/formregistro.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, 'mensaje': 'Error'.format(ex)})

            elif action == 'buscarpersona':
                try:
                    q = request.GET['q'].upper().strip()
                    s = q.split(" ")
                    tipo=request.GET.get('tipo','')
                    if tipo == 'inscripcion':
                        inscripcion=Inscripcion.objects.values_list('persona__id', flat=True).filter(status=True)
                        qspersona = Persona.objects.filter(status=True, id__in=inscripcion).order_by('apellido1')
                    elif tipo == 'externos':
                        exter = Externo.objects.values_list('persona__id', flat=True).filter(status=True)
                        qspersona = Persona.objects.filter(status=True, id__in=exter).order_by('apellido1')
                    else:
                        qspersona = Persona.objects.filter(status=True).order_by('apellido1')

                    if len(s) == 1:
                        qspersona = qspersona.filter((Q(nombres__icontains=q) | Q(apellido1__icontains=q) |
                                                      Q(cedula__icontains=q) | Q(apellido2__icontains=q) |
                                                      Q(usuario__username__icontains=q)))[:15]
                    elif len(s) == 2:
                        qspersona = qspersona.filter((Q(apellido1__contains=s[0]) & Q(apellido2__contains=s[1])) |
                                                     (Q(nombres__icontains=s[0]) & Q(nombres__icontains=s[1])) |
                                                     (Q(nombres__icontains=s[0]) & Q(apellido1__contains=s[1]))).distinct()[:15]
                    else:
                        qspersona = qspersona.filter(
                            (Q(nombres__contains=s[0]) & Q(apellido1__contains=s[1]) & Q(apellido2__contains=s[2])) |
                            (Q(nombres__contains=s[0]) & Q(nombres__contains=s[1]) & Q(
                                apellido1__contains=s[2]))).distinct()[:15]
                    resp = [{'id': qs.pk, 'text': f"{qs.nombre_completo_inverso()}",
                             'documento': qs.documento(),
                             'foto': qs.get_foto()} for qs in qspersona]
                    return HttpResponse(json.dumps({'status': True, 'results': resp}))
                except Exception as ex:
                    pass

            elif action == 'listinscripcion':
                try:
                    lista = []
                    idpersona = int(request.GET['idpersona'])
                    qspersona=Persona.objects.values('nombres','apellido1','apellido2','telefono','telefono_conv','email','emailinst','cedula','pasaporte').get(id=idpersona)
                    inscripciones = Inscripcion.objects.filter(status=True, persona_id=idpersona)
                    for p in inscripciones:
                        activa='Inscripción inactiva' if not p.perfil_inscripcion() else 'Inscripción activa'
                        text = "{} ({})".format(str(p.carrera), str(activa))
                        lista.append([p.id, text])
                    return JsonResponse({'result': 'ok', 'lista': lista, 'datos':qspersona})
                except Exception as e:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

            elif action == 'buscarprofesor':
                try:
                    q = request.GET['q'].upper().strip()
                    s = q.split(" ")
                    qspersona = Profesor.objects.filter(status=True).order_by('persona__apellido1')
                    if len(s) == 1:
                        qspersona = qspersona.filter((Q(persona__nombres__icontains=q) | Q(persona__apellido1__icontains=q) |
                                                      Q(persona__cedula__icontains=q) | Q(persona__apellido2__icontains=q) |
                                                      Q(persona__usuario__username__icontains=q)))[:15]
                    elif len(s) == 2:
                        qspersona = qspersona.filter((Q(persona__apellido1__contains=s[0]) & Q(persona__apellido2__contains=s[1])) |
                                                     (Q(persona__nombres__icontains=s[0]) & Q(persona__nombres__icontains=s[1])) |
                                                     (Q(persona__nombres__icontains=s[0]) & Q(persona__apellido1__contains=s[1]))).distinct()[:15]
                    else:
                        qspersona = qspersona.filter(
                            (Q(persona__nombres__contains=s[0]) & Q(persona__apellido1__contains=s[1]) & Q(persona__apellido2__contains=s[2])) |
                            (Q(persona__nombres__contains=s[0]) & Q(persona__nombres__contains=s[1]) & Q(persona__apellido1__contains=s[2]))).distinct()[:15]
                    resp = [{'id': qs.pk, 'text': f"{qs.persona.nombre_completo_inverso()}",
                             'documento': qs.persona.documento(),
                             'foto': qs.persona.get_foto()} for qs in qspersona]
                    return HttpResponse(json.dumps({'status': True, 'results': resp}))
                except Exception as ex:
                    pass

            elif action == 'buscarlibro':
                try:
                    q = request.GET['q'].upper().strip()
                    s = q.split(" ")
                    filtros = Q(status=True) & ((Q(nombre__icontains=q) | Q(codigokoha__icontains=q) |
                                                  Q(codigoisbn__icontains=q) | Q(titulo__icontains=q) |
                                                  Q(autor__icontains=q)| Q(areaconocimiento__nombre__icontains=q)))
                    qslibro = LibroKohaProgramaAnaliticoAsignatura.objects.filter(filtros).order_by('nombre')[:15]

                    resp = [{'id': qs.pk, 'text': f"{qs.nombre}",
                             'codigo': f"{qs.codigokoha}"} for qs in qslibro]
                    return HttpResponse(json.dumps({'status': True, 'results': resp}))
                except Exception as ex:
                    pass

            elif action == 'infoinscripcion':
                try:
                    lista = []
                    idinscripcion = int(request.GET['idinscripcion'])
                    inscripcion = Inscripcion.objects.get(id=idinscripcion)
                    matricula=inscripcion.matricula_periodo_gimnacio()
                    estadomatricula=''
                    gratuidad=''
                    bloqueada=''
                    gruposocieco=''
                    titulacion=''
                    egresado=''
                    nivel=str(matricula.nivelmalla)
                    if matricula.retiradomatricula:
                        estadomatricula='Retirado'
                    else:
                        estadomatricula=matricula.get_estado_matricula_display()
                        gratuidad=matricula.estadogratuidad()
                    if matricula.bloqueomatricula:
                        bloqueada='Matricula Bloqueada'
                    if matricula.matriculagruposocioeconomico():
                        gruposocieco=matricula.matriculagruposocioeconomico().nombre
                    if inscripcion.proceso_titulacion():
                        titulacion=inscripcion.proceso_titulacion().alternativa.tipotitulacion.get_tipo_display()

                    if inscripcion.graduado() :
                        egresado="GRADUADO"
                    elif inscripcion.egresado():
                        egresado="EGRESADO"

                    periodo_ins=matricula.nivel.periodo.nombre
                    carrera=str(inscripcion.carrera)
                    diccionario={'nivel':nivel,
                                 'estadomatricula':estadomatricula,
                                 'gratuidad':gratuidad,
                                 'bloqueada':bloqueada,
                                 'gruposocieco':gruposocieco,
                                 'titulacion':titulacion,
                                 'periodo':periodo_ins,
                                 'carrera':carrera,
                                 'egresado':egresado}
                    return JsonResponse({'result': 'ok', 'datos': diccionario})
                except Exception as e:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

            #Docentes y administrativos
            elif action == 'addregistrodocente':
                try:
                    form=RegistroDocenteAdministrativoForm()
                    form.cargar()
                    data['form']=form
                    template = get_template("adm_crai/modal/formregistrodocente.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, 'mensaje': 'Error'.format(ex)})

            elif action == 'buscaradmin':
                try:
                    q = request.GET['q'].upper().strip()
                    s = q.split(" ")
                    qspersona = DistributivoPersona.objects.filter(status=True).order_by('persona__apellido1')
                    # if len(ss) == 1:
                    #     administrativos = DistributivoPersona.objects.filter(Q(estadopuesto__id=PUESTO_ACTIVO_ID) & (
                    #                 Q(regimenlaboral__id=1) | Q(regimenlaboral__id=2) | Q(regimenlaboral__id=4)) &
                    #                 (Q(persona__nombres__icontains=search) |
                    #                  Q(persona__apellido1__icontains=search) |
                    #                  Q(persona__apellido2__icontains=search) |
                    #                  Q(persona__cedula__icontains=search) |
                    #                  Q(persona__pasaporte__icontains=search) |
                    #                  Q(denominacionpuesto__descripcion__icontains=search))).exclude(id__in=listadist)
                    # else:
                    #     administrativos = DistributivoPersona.objects.filter( Q(estadopuesto__id=PUESTO_ACTIVO_ID) & (Q(regimenlaboral__id=1) | Q(regimenlaboral__id=2) | Q(regimenlaboral__id=4)) &
                    #         (Q(persona__nombres__icontains=ss[0]) & Q( persona__nombres__icontains=ss[1]) |
                    #          Q(persona__apellido1__icontains=ss[0]) & Q(persona__apellido2__icontains=ss[1]) |
                    #          Q(denominacionpuesto__descripcion__icontains=ss[0]) & Q( denominacionpuesto__descripcion__icontains=ss[1]))).exclude(id__in=listadist)
                    if len(s) == 1:
                        qspersona = qspersona.filter((Q(persona__nombres__icontains=q) | Q(persona__apellido1__icontains=q) |
                                                      Q(persona__cedula__icontains=q) | Q(persona__apellido2__icontains=q) |
                                                      Q(persona__usuario__username__icontains=q)))[:15]
                    elif len(s) == 2:
                        qspersona = qspersona.filter((Q(persona__apellido1__contains=s[0]) & Q(persona__apellido2__contains=s[1])) |
                                                     (Q(persona__nombres__icontains=s[0]) & Q(persona__nombres__icontains=s[1])) |
                                                     (Q(persona__nombres__icontains=s[0]) & Q(persona__apellido1__contains=s[1]))).distinct()[:15]
                    else:
                        qspersona = qspersona.filter(
                            (Q(persona__nombres__contains=s[0]) & Q(persona__apellido1__contains=s[1]) & Q(persona__apellido2__contains=s[2])) |
                            (Q(persona__nombres__contains=s[0]) & Q(persona__nombres__contains=s[1]) & Q(persona__apellido1__contains=s[2]))).distinct()[:15]
                    resp = [{'id': qs.pk, 'text': f"{qs.persona.nombre_completo_inverso()}",
                             'documento': qs.persona.documento(),
                             'foto': qs.persona.get_foto()} for qs in qspersona]
                    return HttpResponse(json.dumps({'status': True, 'results': resp}))
                except Exception as ex:
                    pass

            # Externos
            elif action == 'addregistroexterno':
                try:
                    form=RegistroExternoForm()
                    TIPO_REGISTRO=((1, 'Visita profesor'),(2, 'Visita biblioteca'))
                    form.cargar()
                    form.fields['tiporegistro'].choices=TIPO_REGISTRO
                    data['form']=form
                    template = get_template("adm_crai/modal/formregistroexterno.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, 'mensaje': 'Error'.format(ex)})

            # Reporteria
            elif action == 'reporteria':
                try:
                    if not 'generar' in request.GET:
                        data['servicios']=TipoServicioCrai.objects.filter(status=True)
                        template = get_template("adm_crai/modal/formreporteria.html")
                        return JsonResponse({"result": True, 'data': template.render(data)})
                    else:
                        titulo='Generación de reporte de acceso al crai en proceso.'
                        noti = Notificacion(cuerpo='Reporte de acceso al crai en progreso',
                                            titulo=titulo, destinatario=persona,
                                            url='',
                                            prioridad=1, app_label='SGA',
                                            fecha_hora_visible=datetime.now() + timedelta(days=1), tipo=2,
                                            en_proceso=True)
                        noti.save(request)
                        reporte_acceso_crai_background(request=request, data=data, notif=noti.pk).start()

                except Exception as ex:
                    return JsonResponse({"result": False, 'mensaje': 'Error'.format(ex)})

            elif action == 'cubiculos':
                try:
                    data['title'] = u'Cubículos'
                    search, url_vars, filtros, piso = request.GET.get('search', ''), \
                                                      f'&action={action}', \
                                                      Q(status=True), \
                                                      request.GET.get('piso', '')
                    if search:
                        data['s'] = search
                        filtros = filtros & (Q(nombre__icontains=search) |Q(numero__icontains=search))
                        url_vars += '&s=' + search
                    if piso:
                        data['piso'] = piso = int(piso)
                        filtros = filtros & Q(piso=piso)
                        url_vars += f'&piso={piso}'
                    cubiculos = CubiculoCrai.objects.filter(filtros).order_by('numero')
                    paging = MiPaginador(cubiculos, 10)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['url_vars'] = url_vars
                    data['listado'] = page.object_list
                    data['pisos'] = PisosChoice
                    data['hora'] = datetime.now().time().strftime("%H:%M")
                    request.session['viewactivo'] = 5
                    return render(request, "adm_crai/cubiculos.html", data)
                except Exception as ex:
                    pass

            elif action == 'addcubiculo':
                try:
                    form = CubiculoForm()
                    data['form'] = form
                    template = get_template("ajaxformmodal.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, 'mensaje': 'Error'.format(ex)})

            elif action == 'editcubiculo':
                try:
                    instancia=CubiculoCrai.objects.get(id=encrypt_id(request.GET['id']))
                    form = CubiculoForm(initial=model_to_dict(instancia))
                    data['id']=instancia.id
                    data['form'] = form
                    template = get_template("ajaxformmodal.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, 'mensaje': 'Error'.format(ex)})
            return HttpResponseRedirect(request.path)
        else:
            try:
                data['title'] = u'Estudiantes'
                search, url_vars, filtros, finicio, ffin, tipo, tipo_b, cubiculo = request.GET.get('search', ''), '', \
                                                                         Q(status=True, regimenlaboral_id__isnull=True, inscripcion_id__isnull=False, horafin__isnull=True), \
                                                                         request.GET.get('finicio', ''), \
                                                                         request.GET.get('ffin', ''), \
                                                                         request.GET.get('tipo_ser', ''), \
                                                                         int(request.GET.get('tipo_b', 1)), \
                                                                         request.GET.get('cubiculo', '')
                template="adm_crai/viewestudiantes.html"
                tiposervicio=TipoServicioCrai.objects.filter(status=True)
                if tipo_b == 2:
                    filtros=Q(status=True, regimenlaboral_id__in=[1,2,4],persona__isnull=False, horafin__isnull=True)
                    data['title'] = u'Docentes y administrativos'
                    tiposervicio = tiposervicio.filter(id__in=[1, 8])
                    data['cubiculos'] = CubiculoCrai.objects.filter(status=True)
                    template = "adm_crai/viewdocentesadmin.html"
                elif tipo_b == 3:
                    exter = Externo.objects.values_list('persona__id', flat=True).filter(status=True)
                    filtros=Q(status=True, inscripcion__isnull=True, regimenlaboral_id__isnull=True, persona_id__in=exter, horafin__isnull=True)
                    data['title'] = u'Externos'
                    template="adm_crai/viewexternos.html"
                elif tipo_b == 4:
                    filtros=Q(status=True)
                    data['title'] = u'Historial'
                    template = "adm_crai/viewhistorial.html"
                url_vars += f'&tipo_b={tipo_b}'

                if finicio:
                    data['finicio'] = finicio
                    filtros = filtros & Q(fecha__gte=finicio)
                    url_vars += '&finicio=' + finicio
                elif not tipo_b == 4:
                    filtros = filtros & Q(fecha__gte=datetime.now().date())
                if ffin:
                    data['ffin'] = ffin
                    filtros = filtros & Q(fecha__lte=ffin)
                    url_vars += '&ffin=' + ffin
                if tipo:
                    data['tipo'] = idtipo = int(tipo)
                    filtros = filtros & Q(tiposerviciocrai_id=idtipo)
                    url_vars += '&tipo=' + tipo
                if search:
                    data['s'] = search
                    ss = search.split(' ')
                    if len(ss) == 1:
                        filtros = filtros & (Q(persona__nombres__icontains=search) |
                                             Q(persona__apellido1__icontains=search) |
                                             Q(persona__apellido2__icontains=search) |
                                             Q(persona__cedula__icontains=search))
                        # .exclude(horafin__isnull=False)
                    else:
                        filtros = filtros & (Q(persona__nombres__icontains=ss[0])|
                                             Q(persona__apellido1__icontains=ss[0]) & Q(persona__apellido2__icontains=ss[1]))
                    url_vars +='&s='+search

                if cubiculo:
                    data['cubiculo'] = cubiculo = int(cubiculo)
                    filtros = filtros & Q(cubiculo_id=cubiculo)
                    url_vars += f'&cubiculo={cubiculo}'

                visitas = RegistrarIngresoCrai.objects.filter(filtros).order_by('-fecha','-horainicio')
                paging = MiPaginador(visitas, 10)
                p = 1
                try:
                    paginasesion = 1
                    if 'paginador' in request.session:
                        paginasesion = int(request.session['paginador'])
                    if 'page' in request.GET:
                        p = int(request.GET['page'])
                    else:
                        p = paginasesion
                    try:
                        page = paging.page(p)
                    except:
                        p = 1
                    page = paging.page(p)
                except:
                    page = paging.page(p)
                request.session['paginador'] = p
                data['paging'] = paging
                data['rangospaging'] = paging.rangos_paginado(p)
                data['page'] = page
                data['url_vars']=url_vars
                data['visitas'] = page.object_list
                data['tipo_servicio'] = tiposervicio
                data['hora'] = datetime.now().time().strftime("%H:%M")
                request.session['viewactivo'] = tipo_b
                return render(request, template, data)
            except Exception as ex:
                pass


