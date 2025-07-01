# -*- coding: UTF-8 -*-
import json
import os
import sys
import zipfile
import random

from openpyxl import workbook as openxl
from openpyxl.styles import Font as openxlFont
from openpyxl.styles.alignment import Alignment as alin
from openpyxl.chart import ScatterChart, Reference, Series, PieChart, BarChart
from datetime import timedelta, datetime

from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.forms import model_to_dict
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.template.loader import get_template
from django.shortcuts import render, redirect
from django.contrib import messages

from certi.models import Certificado, CertificadoAsistenteCertificadora
from decorators import secure_module, last_access
from homologa.models import PeriodoHomologacion, ResponsableHomologacion, SolicitudEstudianteHomologacion, \
    RequisitosHomologacion, RequisitoPeriodoHomologacion, ESTADO_SOLICITUD, ESTADO_VALIDACION, \
    ESTADO_VALIDACION_REQUISITOS, DocumentosSolicitudHomologacion, SeguimientoRevision
from homologa.forms import PeriodoHomologacionForm, ResponsableHomologacionForm, RequisitosHomologacionForm, \
    RequisitosPeriodoForm, ValidarRequisitosForm, ValidarDirectorForm, SubirResolucionForm
from settings import SITE_STORAGE
from sga.commonviews import adduserdata
from sga.funciones import MiPaginador, log, generar_nombre, notificacion, remover_caracteres_especiales_unicode
from django.db.models import Q

from sga.models import Persona, Malla, CUENTAS_CORREOS, CoordinadorCarrera, Coordinacion, Carrera, Periodo, Materia
from sga.tasks import send_html_mail
from sga.templatetags.sga_extras import encrypt


@login_required(redirect_field_name='ret', login_url='/loginsga')
@secure_module
@last_access
# @transaction.atomic()
def view(request):
    data = {}
    adduserdata(request, data)
    persona = request.session['persona']
    periodo = request.session['periodo']
    hoy=datetime.now()
    certificado = Certificado.objects.filter(status=True).first()
    data['responsable_ga']=responsable_ga=ResponsableHomologacion.objects.filter(status=True,persona=persona).first()
    data['es_director']=es_director=CoordinadorCarrera.objects.filter(status=True, persona=persona, tipo=3, periodo=periodo).last()
    data['es_asistente']=es_asistente=ResponsableHomologacion.objects.filter(status=True, persona=persona, estado=True, rol=2).last()
    if not responsable_ga and not es_director and not es_asistente and not request.user.is_superuser and not request.user.has_perm('sga.puede_revisar_total_solicitudes_homologacion'):
        return HttpResponseRedirect("/?info=Usted no tiene acceso a este módulo.")
    if request.method == 'POST':
        action = request.POST['action']

        # Periodos de homologación
        if action == 'addperiodo':
            with transaction.atomic():
                try:
                    form = PeriodoHomologacionForm(request.POST)
                    if form.is_valid():
                        instancia = PeriodoHomologacion(periodo=form.cleaned_data['periodo'],
                                                        fechaapertura=form.cleaned_data['fechaapertura'],
                                                        fechacierre =form.cleaned_data['fechacierre'],
                                                        fechainiciorecepciondocumentos=form.cleaned_data['fechainiciorecepciondocumentos'],
                                                        fechacierrerecepciondocumentos=form.cleaned_data['fechacierrerecepciondocumentos'],
                                                        fechainiciorevisiongacademica=form.cleaned_data['fechainiciorevisiongacademica'],
                                                        fechacierrerevisiongacademica=form.cleaned_data['fechacierrerevisiongacademica'],
                                                        fechainiciovaldirector=form.cleaned_data['fechainiciovaldirector'],
                                                        fechacierrevaldirector=form.cleaned_data['fechacierrevaldirector'],
                                                        fechainicioremitiraprobados=form.cleaned_data['fechainicioremitiraprobados'],
                                                        fechacierreremitiraprobados=form.cleaned_data['fechacierreremitiraprobados'],
                                                        publico=form.cleaned_data['publico'],
                                                        numdias=form.cleaned_data['numdias'],
                                                        motivo=form.cleaned_data['motivo'])
                        instancia.save(request)
                        if instancia.publico:
                            periodos = PeriodoHomologacion.objects.filter(status=True, publico=True).exclude(pk=instancia.id)
                            for p in periodos:
                                p.publico = False
                                p.save(request)
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()],
                                             "mensaje": "Error en el formulario"})
                    messages.success(request, 'Guardado con exito')
                    log(u'Adiciono actividad de ayudantias de investigacion: %s' % instancia, request, "add")
                    return JsonResponse({"result": False}, safe=False)
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "mensaje": 'Error: {}'.format(ex)}, safe=False)

        if action == 'editperiodo':
            with transaction.atomic():
                try:
                    filtro = PeriodoHomologacion.objects.get(id=int(encrypt(request.POST['id'])))
                    form = PeriodoHomologacionForm(request.POST)
                    if form.is_valid():
                        filtro.periodo = form.cleaned_data['periodo']
                        filtro.fechaapertura = form.cleaned_data['fechaapertura']
                        filtro.fechacierre = form.cleaned_data['fechacierre']
                        filtro.fechainiciorecepciondocumentos = form.cleaned_data['fechainiciorecepciondocumentos']
                        filtro.fechacierrerecepciondocumentos = form.cleaned_data['fechacierrerecepciondocumentos']
                        filtro.fechainiciorevisiongacademica = form.cleaned_data['fechainiciorevisiongacademica']
                        filtro.fechacierrerevisiongacademica = form.cleaned_data['fechacierrerevisiongacademica']
                        filtro.fechainiciovaldirector = form.cleaned_data['fechainiciovaldirector']
                        filtro.fechacierrevaldirector = form.cleaned_data['fechacierrevaldirector']
                        filtro.fechainicioremitiraprobados = form.cleaned_data['fechainicioremitiraprobados']
                        filtro.fechacierreremitiraprobados = form.cleaned_data['fechacierreremitiraprobados']
                        filtro.publico = form.cleaned_data['publico']
                        filtro.numdias = form.cleaned_data['numdias']
                        filtro.motivo = form.cleaned_data['motivo']
                        filtro.save(request)
                        if filtro.publico:
                            periodos = PeriodoHomologacion.objects.filter(status=True, publico=True).exclude(pk=filtro.id)
                            for p in periodos:
                                p.publico = False
                                p.save(request)
                        log(u'Edito periodo de homologación: %s' % filtro, request, "edit")
                        return JsonResponse({"result": False}, safe=False)
                    else:
                        return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()],
                                             "mensaje": "Error en el formulario"})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "mensaje": 'Error: {}'.format(ex)}, safe=False)

        if action == 'delperiodo':
            with transaction.atomic():
                try:
                    instancia = PeriodoHomologacion.objects.get(pk=int(encrypt(request.POST['id'])))
                    instancia.status = False
                    instancia.save(request)
                    log(u'Elimino periodo de homologación: %s' % instancia, request, "del")
                    res_json = {"error": False}
                except Exception as ex:
                    res_json = {'error': True, "message": "Error: {}".format(ex)}
                return JsonResponse(res_json, safe=False)

        if action == 'publicarperiodo':
            with transaction.atomic():
                try:
                    publico = eval(request.POST['val'].capitalize())
                    if publico:
                        periodos = PeriodoHomologacion.objects.filter(status=True, publico=True).exclude(pk=int(request.POST['id']))
                        for p in periodos:
                            p.publico = False
                            p.save(request)
                    registro = PeriodoHomologacion.objects.get(pk=int(request.POST['id']))
                    registro.publico = publico
                    registro.save(request)
                    log(u'Publicar periodo de homologación: %s (%s)' % (registro, registro.publico), request,"edit")
                    return JsonResponse({"result": True, 'mensaje': 'Cambios guardados'})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": False})

        # Responsables
        if action == 'addresponsable':
            with transaction.atomic():
                try:
                    form = ResponsableHomologacionForm(request.POST)
                    if form.is_valid():
                        instancia = ResponsableHomologacion(persona=form.cleaned_data['persona'],
                                                            rol=form.cleaned_data['rol'],
                                                            estado =form.cleaned_data['estado'],
                                                            carrera=form.cleaned_data.get('carrera', None)
                                                            )
                        instancia.save(request)
                        for c in form.cleaned_data['coordinaciones']:
                            instancia.coordinaciones.add(c)
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()],
                                             "mensaje": "Error en el formulario"})
                    messages.success(request, 'Guardado con exito')
                    log(u'Adiciono responsable: %s' % instancia, request, "add")
                    return JsonResponse({"result": False}, safe=False)
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "mensaje": 'Error: {}'.format(ex)}, safe=False)

        if action == 'editresponsable':
            with transaction.atomic():
                try:
                    filtro = ResponsableHomologacion.objects.get(id=int(encrypt(request.POST['id'])))
                    form = ResponsableHomologacionForm(request.POST, instancia=filtro)
                    if form.is_valid():
                        filtro.persona= form.cleaned_data['persona']
                        filtro.rol = form.cleaned_data['rol']
                        filtro.estado = form.cleaned_data['estado']
                        filtro.carrera = form.cleaned_data.get('carrera', None)
                        filtro.save(request)
                        filtro.coordinaciones.clear()
                        for c in form.cleaned_data['coordinaciones']:
                            filtro.coordinaciones.add(c)
                        log(u'Edito responsable: %s' % filtro, request, "edit")
                        return JsonResponse({"result": False}, safe=False)
                    else:
                        return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()],
                                             "mensaje": "Error en el formulario"})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "mensaje": 'Error: {}'.format(ex)}, safe=False)

        if action == 'delresponsable':
            with transaction.atomic():
                try:
                    instancia = ResponsableHomologacion.objects.get(pk=int(encrypt(request.POST['id'])))
                    instancia.status = False
                    instancia.save(request)
                    log(u'Elimino responsable: %s' % instancia, request, "del")
                    res_json = {"error": False}
                except Exception as ex:
                    res_json = {'error': True, "message": "Error: {}".format(ex)}
                return JsonResponse(res_json, safe=False)

        if action == 'activarresponsable':
            with transaction.atomic():
                try:
                    estado = eval(request.POST['val'].capitalize())
                    registro = ResponsableHomologacion.objects.get(pk=int(request.POST['id']))
                    registro.estado = estado
                    registro.save(request)
                    if estado and registro.rol == 2:
                        ids_coordinaciones = registro.coordinaciones.all().values_list('id', flat=True)
                        asistentes = ResponsableHomologacion.objects.filter(status=True, estado=True, coordinaciones__id__in=ids_coordinaciones).exclude(id=registro.id)
                        if asistentes:
                            asistentes.update(estado=False)
                    log(u'Edito responsable activo: %s (%s)' % (registro, registro.estado), request,"edit")
                    return JsonResponse({"result": True, 'mensaje': 'Cambios guardados'})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": False})

        # Requisitos
        if action == 'addrequisito':
            with transaction.atomic():
                try:
                    form = RequisitosHomologacionForm(request.POST)
                    if form.is_valid():
                        instancia = RequisitosHomologacion(nombre=form.cleaned_data['nombre'],
                                                            leyenda=form.cleaned_data['leyenda'])
                        instancia.save(request)
                        if 'archivo' in request.FILES:
                            newfile = request.FILES['archivo']
                            extension = newfile._name.split('.')
                            tam = len(extension)
                            exte = extension[tam - 1]
                            if newfile.size > 2194304:
                                raise NameError('El tamaño del archivo es mayor a 2 Mb.')
                            if not exte.lower() in ['pdf']:
                                raise NameError('Solo archivos .pdf')
                            newfile._name = generar_nombre(f'requisito_{instancia.id}_',newfile._name)
                            instancia.archivo = newfile
                            instancia.save(request)
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()],"mensaje": "Error en el formulario"})
                    messages.success(request, 'Guardado con exito')
                    log(u'Adiciono requisito: %s' % instancia, request, "add")
                    return JsonResponse({"result": False}, safe=False)
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "mensaje": 'Error: {}'.format(ex)}, safe=False)

        if action == 'editrequisito':
            with transaction.atomic():
                try:
                    filtro = RequisitosHomologacion.objects.get(id=int(encrypt(request.POST['id'])))
                    form = RequisitosHomologacionForm(request.POST)
                    if form.is_valid():
                        filtro.nombre = form.cleaned_data['nombre']
                        filtro.leyenda = form.cleaned_data['leyenda']
                        filtro.save(request)
                        if 'archivo' in request.FILES:
                            newfile = request.FILES['archivo']
                            extension = newfile._name.split('.')
                            tam = len(extension)
                            exte = extension[tam - 1]
                            if newfile.size > 2194304:
                                raise NameError('Error, el tamaño del archivo es mayor a 2 Mb.')
                            if not exte.lower() in ['pdf']:
                                raise NameError('Error, solo archivos .pdf')
                            newfile._name = generar_nombre(f'requisito_{filtro.id}_',newfile._name)
                            filtro.archivo = newfile
                            filtro.save(request)
                        log(u'Edito requisito: %s' % filtro, request, "edit")
                        return JsonResponse({"result": False}, safe=False)
                    else:
                        return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()], "mensaje": "Error en el formulario"})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "mensaje": 'Error: {}'.format(ex)}, safe=False)

        if action == 'delrequisitos':
            with transaction.atomic():
                try:
                    instancia = RequisitosHomologacion.objects.get(pk=int(encrypt(request.POST['id'])))
                    instancia.status = False
                    instancia.save(request)
                    log(u'Elimino requisito de homologacion: %s' % instancia, request, "del")
                    res_json = {"error": False}
                except Exception as ex:
                    res_json = {'error': True, "message": "Error: {}".format(ex)}
                return JsonResponse(res_json, safe=False)

        # Requisitos Periodo
        if action == 'addrequisitoperiodo':
            with transaction.atomic():
                try:
                    idperiodo = int(encrypt(request.POST['id']))
                    form = RequisitosPeriodoForm(request.POST)
                    if form.is_valid() and form.validador(idperiodo):
                        requisito = RequisitoPeriodoHomologacion(periodo_h_id=idperiodo,
                                                          requisito=form.cleaned_data['requisito'],
                                                          # opcional=form.cleaned_data['opcional'],
                                                          # archivo=form.cleaned_data['archivo'],
                                                          mostrar=True)
                        requisito.save(request)
                        diccionario = {'id': requisito.id,
                                       'idperiodo': requisito.periodo_h.id,
                                       'requisito': requisito.requisito.nombre.capitalize(),
                                       'opcional': requisito.opcional,
                                       'essilabo': requisito.essilabo,
                                       'multiple': requisito.multiple,
                                       'mostrar': requisito.mostrar,
                                       }
                        log(u'Agrego Requisito a Periodo: %s' % requisito, request, "add")
                        return JsonResponse(
                            {'result': True, 'data_return': True, 'mensaje': u'Guardado con exito',
                             'data': diccionario})
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse(
                            {'result': True, "form": [{k: v[0]} for k, v in form.errors.items()],
                             "mensaje": "Error en el formulario"})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "mensaje": 'Error: {}'.format(ex)}, safe=False)

        if action == 'editrequisitoperiodo':
            with transaction.atomic():
                try:
                    requisito = RequisitoPeriodoHomologacion.objects.get(pk=int(request.POST['id']))
                    if request.POST['name'] == 'mostrar':
                        requisito.mostrar = eval(request.POST['val'].capitalize())
                        requisito.save(request)
                        log(u'Edito estado mostrar de requisito : %s (%s)' % (requisito, requisito.mostrar),request,"editrequisito")

                    if request.POST['name'] == 'opcional':
                        requisito.opcional = eval(request.POST['val'].capitalize())
                        requisito.save(request)
                        log(u'Edito estado opcional de requisito : %s (%s)' % (requisito, requisito.opcional),request,"editrequisito")

                    if request.POST['name'] == 'multiple':
                        requisito.multiple = eval(request.POST['val'].capitalize())
                        requisito.save(request)
                        log(u'Edito estado multiple de requisito : %s (%s)' % (requisito, requisito.multiple),request,"editrequisito")

                    if request.POST['name'] == 'silabo':
                        requisito.essilabo = eval(request.POST['val'].capitalize())
                        requisito.save(request)
                        log(u'Edito estado multiple de requisito : %s (%s)' % (requisito, requisito.essilabo),request,"editrequisito")

                    return JsonResponse({"result": True, 'mensaje': 'Cambios guardados'})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': False, "mensaje": 'Error: {}'.format(ex)}, safe=False)

        if action == 'delrequisitoperiodo':
            with transaction.atomic():
                try:
                    instancia = RequisitoPeriodoHomologacion.objects.get(pk=int(request.POST['id']))
                    instancia.status = False
                    instancia.save(request)
                    log(u'Elimino requisito de periodo: %s' % instancia, request, "del")
                    res_json = {"error": False, "mensaje": 'Registro eliminado'}
                except Exception as ex:
                    res_json = {'error': True, "message": "Error: {}".format(ex)}
                return JsonResponse(res_json, safe=False)

        #Validar Requisitos
        if action == 'validarrequisitos':
            with transaction.atomic():
                try:
                    instance = DocumentosSolicitudHomologacion.objects.get(pk=int(request.POST['id']))
                    form = ValidarRequisitosForm(request.POST)
                    if form.is_valid():
                        instance.estado = int(form.cleaned_data['estado'])
                        instance.observacion = form.cleaned_data['observacion']
                        instance.f_validacion=hoy
                        instance.save(request)
                        mensaje = ''
                        solicitud = instance.solicitud
                        solicitud.fecha_revision_gacademica= hoy
                        enviar_correo=False
                        if solicitud.estado != 3:
                            solicitud.estado = 3
                        # Aprobado todos los requisitos
                        if instance.estado == 1 and solicitud.doc_validacion() == 1:
                            solicitud.revision_gacademica = 1
                            mensaje = 'Su solicitud de homologación de asignaturas fue aprobado por el departamento de Gestión Academica'
                            solicitud.observacion_gacademica = 'Requisitos validados correctamente y la solicitud es trasladada al siguiente paso de validación por parte del director de carrera.'
                            enviar_correo = True
                        #Rechazado todos los requisitos
                        elif instance.estado == 2 and solicitud.doc_validacion() == 2:
                            solicitud.revision_gacademica = 2
                            mensaje = 'Su solicitud de homologación de asignaturas tiene documentos rechazados'
                            solicitud.observacion_gacademica=''
                            enviar_correo = True
                        #Corregir al menos un requisito
                        elif instance.estado == 3 or solicitud.doc_validacion() == 3:
                            solicitud.revision_gacademica = 3
                            if instance.estado == 3:
                                mensaje = 'Su solicitud de homologación de asignaturas tiene documentos por corregir'
                                enviar_correo=True
                        # Retornar a pendiente
                        elif solicitud.doc_validacion() == 0:
                            solicitud.revision_gacademica = 0
                        solicitud.save(request)

                        if enviar_correo:
                            titulo = u"Validación de documentos de solicitud de homologación ({})".format(instance.get_estado_display())
                            notificacion(titulo, mensaje, solicitud.inscripcion.persona, None, u'/alu_homologacion?action=verproceso&id={}'.format(encrypt(solicitud.id)),solicitud.pk, 1, 'sga', SolicitudEstudianteHomologacion, request)
                            lista_email = solicitud.inscripcion.persona.lista_emails()
                            # lista_email = ['jguachuns@unemi.edu.ec', ]
                            datos_email = {'sistema': request.session['nombresistema'],
                                           'fecha': datetime.now().date(),
                                           'hora': datetime.now().time(),
                                           'documento': instance,
                                           'persona': solicitud.inscripcion.persona,
                                           'mensaje': mensaje}
                            template = "emails/notificacion_homologaasignatura.html"
                            send_html_mail(titulo, template, datos_email, lista_email, [], [],cuenta=CUENTAS_CORREOS[0][1])

                        # Guardar seguimiento de validación
                        seguimiento = SeguimientoRevision(revisor=persona,
                                                         documento=instance, estado_doc=instance.estado,
                                                         solictud=solicitud, estado=solicitud.estado,
                                                         rutaarchivo=instance.archivo.url,
                                                         observacion=instance.observacion)
                        seguimiento.save(request)
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()],
                                             "mensaje": "Error en el formulario"})
                    diccionario={'id':instance.id,
                                 'observacion':instance.observacion,
                                 'estado':instance.get_estado_display(),
                                 'idestado':instance.estado,
                                 'color':instance.color_estado(),
                                 'estado_s':solicitud.get_revision_gacademica_display(),
                                 'color_s':solicitud.color_validacion_gacademico()}
                    log(u'Valido documento de solicitud: %s' % instance, request, "edit")
                    return JsonResponse({'result': True,'data_return':True, 'mensaje': u'Guardado con exito', 'data':diccionario}, safe=False)
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": True, "mensaje": "Intentelo más tarde.".format(str(ex))}, safe=False)

        if action == 'rechazarsolicitud':
            with transaction.atomic():
                try:
                    if not request.POST['observacion']:
                        raise NameError('Por favor escriba el motivo del rechazo de solicitud')
                    instance = SolicitudEstudianteHomologacion.objects.get(pk=int(encrypt(request.POST['id'])))
                    instance.estado = 2
                    instance.revision_gacademica = 2
                    instance.observacion_gacademica = request.POST['observacion']
                    instance.save(request)
                    titulo = u"Solicitud de homologación de asignaturas ({})".format(instance.get_estado_display())
                    mensaje= u"Su solicitud para homologación de asignaturas fue rechazada"
                    notificacion(titulo, mensaje,instance.inscripcion.persona, None, f'/alu_homologacion?action=verproceso&id={encrypt(instance.id)}',instance.pk, 1, 'sga', DocumentosSolicitudHomologacion, request)
                    lista_email = instance.inscripcion.persona.lista_emails()
                    # lista_email = ['jguachuns@unemi.edu.ec', ]
                    datos_email = {'sistema': request.session['nombresistema'],
                                   'fecha': datetime.now().date(),
                                   'hora': datetime.now().time(),
                                   'documento': instance,
                                   'persona': instance.inscripcion.persona,
                                   'mensaje': mensaje}
                    template = "emails/notificacion_homologaasignatura.html"
                    send_html_mail(titulo, template, datos_email, lista_email, [], [],cuenta=CUENTAS_CORREOS[0][1])

                    # Guardar seguimiento de validación
                    seguimiento = SeguimientoRevision(revisor=persona,
                                                      solictud=instance, estado=instance.estado,
                                                      observacion=instance.observacion)
                    seguimiento.save(request)
                    log(u'Rechazo soliditud de homologación: %s' % instance, request, "edit")
                    messages.success(request, 'Solicitud rechazada')
                    return JsonResponse({'result': True,'mensaje': u'Solicitud rechazada'}, safe=False)
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": False, "mensaje": "Intentelo más tarde.".format(str(ex))}, safe=False)

        if action == 'validar_director':
            with transaction.atomic():
                try:
                    instance = SolicitudEstudianteHomologacion.objects.get(pk=int(encrypt(request.POST['id'])))
                    form = ValidarDirectorForm(request.POST, request.FILES)
                    if form.is_valid():
                        newfile = request.FILES['archivo']
                        newfile._name = generar_nombre(f'resolucion_director_{persona.id}_{instance.id}_', newfile._name)
                        instance.persona_director = persona
                        instance.estado = int(form.cleaned_data['revision_director'])
                        instance.revision_director = int(form.cleaned_data['revision_director'])
                        instance.fecha_revision_director=hoy
                        instance.observacion_director= form.cleaned_data['observacion_director']
                        instance.archivoresoluciondirector = newfile
                        instance.save(request)

                        titulo = u"Validación de solicitud de homologación de asignaturas ({})".format(instance.get_estado_display())
                        mensaje = 'Su solicitud de homologación fue validado por el director de carrera'
                        notificacion(titulo, mensaje, instance.inscripcion.persona, None,
                                     u'/alu_homologacion?action=verproceso&id={}'.format(encrypt(instance.id)),
                                     instance.pk, 1, 'sga', SolicitudEstudianteHomologacion, request)
                        lista_email = instance.inscripcion.persona.lista_emails()
                        # lista_email = ['jguachuns@unemi.edu.ec', ]
                        datos_email = {'sistema': request.session['nombresistema'],
                                       'fecha': datetime.now().date(),
                                       'hora': datetime.now().time(),
                                       'solicitud': instance,
                                       'persona': instance.inscripcion.persona,
                                       'observacion':instance.observacion_director,
                                       'mensaje': mensaje}
                        template = "emails/notificacion_homologaasignatura.html"
                        send_html_mail(titulo, template, datos_email, lista_email, [], [],cuenta=CUENTAS_CORREOS[0][1])

                        # Guardar seguimiento de validación
                        seguimiento = SeguimientoRevision(revisor=persona,
                                                         solictud=instance, estado=instance.estado,
                                                         rutaarchivo=instance.archivoresoluciondirector.url,
                                                         observacion=instance.observacion_director)
                        seguimiento.save(request)
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()],
                                             "mensaje": "Error en el formulario"})
                    log(u'Valido documento de solicitud: %s' % instance, request, "edit")
                    return JsonResponse({'result': False, 'mensaje': u'Guardado con exito'}, safe=False)
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": True, "mensaje": "Intentelo más tarde.".format(str(ex))}, safe=False)

        if action == 'subirresolucion':
            with transaction.atomic():
                try:
                    instance = SolicitudEstudianteHomologacion.objects.get(pk=int(encrypt(request.POST['id'])))
                    form = SubirResolucionForm(request.POST, request.FILES)
                    if form.is_valid():
                        newfile = request.FILES['archivo_r']
                        newfile._name = generar_nombre(f'resolucion_directivo_{persona.id}_{instance.id}_', newfile._name)
                        imgfile = request.FILES['imgevidencia']
                        imgfile._name = generar_nombre(f'evidencia_{persona.id}_{instance.id}_', imgfile._name)
                        instance.asistente_facultad= persona
                        instance.revision_directivo = int(form.cleaned_data['revision_directivo'])
                        if instance.revision_directivo == 1:
                            instance.estado = 4
                            mensaje = 'Su solicitud de homologación de asignaturas finalizo con éxito, la información fue cargada al sistema'
                        else:
                            instance.estado=2
                            mensaje = 'Su solicitud de homologación de asignaturas fue rechazada por los directivos.'

                        instance.fecha_resolucion_aprobacion=hoy
                        instance.observacion = form.cleaned_data['observacion']
                        instance.archivoresoluciondirectivo = newfile
                        instance.imgevidencia = imgfile
                        instance.save(request)

                        titulo = u"Finalización del proceso de homologación de asignaturas ({})".format(instance.get_estado_display())
                        notificacion(titulo, mensaje, instance.inscripcion.persona, None,
                                     u'/alu_homologacion?action=verproceso&id={}'.format(encrypt(instance.id)),
                                     instance.pk, 1, 'sga', SolicitudEstudianteHomologacion, request)
                        lista_email = instance.inscripcion.persona.lista_emails()
                        # lista_email = ['jguachuns@unemi.edu.ec', ]
                        datos_email = {'sistema': request.session['nombresistema'],
                                       'fecha': datetime.now().date(),
                                       'hora': datetime.now().time(),
                                       'solicitud': instance,
                                       'persona': instance.inscripcion.persona,
                                       'observacion': instance.observacion,
                                       'mensaje': mensaje}
                        template = "emails/notificacion_homologaasignatura.html"
                        send_html_mail(titulo, template, datos_email, lista_email, [], [], cuenta=CUENTAS_CORREOS[0][1])

                        # Guardar seguimiento de validación
                        seguimiento = SeguimientoRevision(revisor=persona,
                                                         solictud=instance, estado=instance.estado,
                                                         rutaarchivo=instance.archivoresoluciondirectivo.url,
                                                         observacion=instance.observacion_director)
                        seguimiento.save(request)
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()],
                                             "mensaje": "Error en el formulario"})
                    log(u'Valido documento de solicitud: %s' % instance, request, "edit")
                    return JsonResponse({'result': False, 'mensaje': u'Guardado con exito'}, safe=False)
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": True, "mensaje": "Intentelo más tarde.".format(str(ex))}, safe=False)

        return JsonResponse({"result": "bad", "mensaje": u"Solicitud Incorrecta."})
    else:
        if 'action' in request.GET:
            data['action'] = action = request.GET['action']

            # Periodos
            if action == 'addperiodo':
                try:
                    form = PeriodoHomologacionForm()
                    form.fields['periodo'].queryset = Periodo.objects.filter(fin__gte=datetime.now().date(), status=True).exclude(tipo_id__in=[1,3])
                    data['form'] = form
                    template = get_template("adm_homologacion/modal/formperiodohomologacion.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            if action == 'editperiodo':
                try:
                    data['id'] = id = int(encrypt(request.GET['id']))
                    data['filtro'] = filtro = PeriodoHomologacion.objects.get(pk=id)
                    form = PeriodoHomologacionForm(initial=model_to_dict(filtro))
                    form.fields['periodo'].queryset = Periodo.objects.filter(Q(fin__gte=datetime.now().date())|Q(id=filtro.periodo.id),status=True).exclude(tipo_id__in=[1,3])
                    data['form'] = form
                    template = get_template("adm_homologacion/modal/formperiodohomologacion.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            # Responsables
            if action == 'responsables':
                try:
                    data['title'] = u'Responsables'
                    search, filtro, url_vars = request.GET.get('s', ''), Q(status=True), f'&action={action}'
                    if search:
                        data['s'] = search
                        s = search.split(' ')
                        if len(s) == 1:
                            filtro = filtro & (Q(persona__nombres__unaccent__icontains=search) |
                                               Q(persona__cedula__icontains=search) |
                                               Q(persona__apellido1__unaccent__icontains=search) |
                                               Q(persona__apellido2__unaccent__icontains=search))
                        if len(s) >= 2:
                            filtro = filtro & (Q(persona__apellido1__unaccent__icontains=s[0]) & Q(persona__apellido2__unaccent__icontains=s[1]))
                        url_vars += '&s=' + search

                    listado = ResponsableHomologacion.objects.filter(filtro).order_by('-id')
                    paging = MiPaginador(listado, 10)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data["url_vars"] = url_vars
                    data['listado'] = page.object_list
                    request.session['viewactivo'] = 2
                    data['listcount'] = len(listado)
                    return render(request, 'adm_homologacion/viewresponsables.html', data)
                except Exception as ex:
                    return render({'result': False, 'mensaje': '{}'.format(ex)})

            if action == 'addresponsable':
                try:
                    form = ResponsableHomologacionForm()
                    form.fields['persona'].queryset = Persona.objects.none()
                    form.fields['carrera'].queryset = Carrera.objects.none()
                    data['form'] = form
                    template = get_template("adm_homologacion/modal/formresponsable.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            if action == 'editresponsable':
                try:
                    data['id'] = id = int(encrypt(request.GET['id']))
                    data['filtro'] = filtro = ResponsableHomologacion.objects.get(pk=id)
                    form = ResponsableHomologacionForm(initial=model_to_dict(filtro))
                    form.fields['persona'].queryset=Persona.objects.filter(id=filtro.persona.id)
                    data['form'] = form
                    template = get_template("adm_homologacion/modal/formresponsable.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            if action == 'buscarpersonas':
                try:
                    idsexcluidas = []
                    idsagregados = request.GET['idsagregados']
                    if idsagregados:
                        idsagregados = idsagregados.split(',')
                        idsexcluidas += [idl for idl in idsagregados]
                    q = request.GET['q'].upper().strip()
                    s = q.split(" ")
                    qspersona = Persona.objects.filter(status=True, administrativo__isnull=False).exclude(id__in=idsexcluidas).order_by('apellido1')
                    if len(s) == 1:
                        qspersona = qspersona.filter((Q(nombres__icontains=q) | Q(apellido1__icontains=q) | Q(
                            cedula__icontains=q) | Q(apellido2__icontains=q) | Q(cedula__contains=q)),
                                                     Q(status=True)).distinct()[:15]
                    elif len(s) == 2:
                        qspersona = qspersona.filter((Q(apellido1__contains=s[0]) & Q(apellido2__contains=s[1])) |
                                                     (Q(nombres__icontains=s[0]) & Q(nombres__icontains=s[1])) |
                                                     (Q(nombres__icontains=s[0]) & Q(apellido1__contains=s[1]))).filter(
                            status=True).distinct()[:15]
                    else:
                        qspersona = qspersona.filter(
                            (Q(nombres__contains=s[0]) & Q(apellido1__contains=s[1]) & Q(apellido2__contains=s[2])) |
                            (Q(nombres__contains=s[0]) & Q(nombres__contains=s[1]) & Q(
                                apellido1__contains=s[2]))).filter(status=True).distinct()[:15]

                    resp = [{'id': qs.pk, 'text': f"{qs.nombre_completo_inverso()}",
                             'documento': qs.documento(),
                             'foto': qs.get_foto()} for qs in qspersona]
                    return HttpResponse(json.dumps({'status': True, 'results': resp}))
                except Exception as ex:
                    pass

            if action == 'buscarcarrera':
                try:
                    coordinacion_id = request.GET['idcoor[]']
                    coordinacion = Coordinacion.objects.filter(pk=coordinacion_id).first()
                    mallascarrera = Materia.objects.filter(nivel__periodo=periodo).values_list(
                        'asignaturamalla__malla__carrera_id', flat=True).distinct()
                    carreras = Carrera.objects.filter(pk__in=mallascarrera,
                                                              status=True,
                                                              coordinacion=coordinacion)
                    carreras_list = [{'id': carrera.id, 'nombre': carrera.nombre} for carrera in carreras]
                    return JsonResponse({'status': True, 'carreras': carreras_list})
                except Exception as ex:
                    pass

            # Requisitos
            if action == 'requisitos':
                try:
                    data['title'] = u'Requisitos'
                    search, filtro, url_vars = request.GET.get('s', ''), Q(status=True), f'&action={action}'
                    if search:
                        data['s'] = search
                        filtro = filtro & (Q(nombre__unaccent__icontains=search))
                        url_vars += '&s=' + search

                    listado = RequisitosHomologacion.objects.filter(filtro).order_by('-id')
                    paging = MiPaginador(listado, 10)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data["url_vars"] = url_vars
                    data['listado'] = page.object_list
                    request.session['viewactivo'] = 3
                    data['listcount'] = len(listado)
                    return render(request, 'adm_homologacion/viewrequisitos.html', data)
                except Exception as ex:
                    return render({'result': False, 'mensaje': '{}'.format(ex)})

            if action == 'addrequisito':
                try:
                    form = RequisitosHomologacionForm()
                    data['form'] = form
                    template = get_template("adm_homologacion/modal/formrequisitos.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            if action == 'editrequisito':
                try:
                    data['id'] = id = int(encrypt(request.GET['id']))
                    data['filtro'] = filtro = RequisitosHomologacion.objects.get(pk=id)
                    form = RequisitosHomologacionForm(initial=model_to_dict(filtro))
                    data['form'] = form
                    template = get_template("adm_homologacion/modal/formrequisitos.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            if action == 'requisitosperiodo':
                try:
                    form = RequisitosPeriodoForm()
                    data['filtro'] = r_periodo = PeriodoHomologacion.objects.get(pk=int(encrypt(request.GET['id'])))
                    data['listado'] = r_periodo.requisitos()
                    data['form'] = form
                    data['requisitos'] = RequisitosHomologacion.objects.filter(status=True)
                    template = get_template("adm_homologacion/modal/formrequisitosperiodo.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            # Solicitudes
            if action == 'solicitudes':
                try:
                    data['title'] = u'Solicitudes de homologación'
                    id=int(encrypt(request.GET['id']))
                    data['periodo_h']=periodo_h=PeriodoHomologacion.objects.get(id=id)
                    es_director1 = CoordinadorCarrera.objects.filter(status=True, persona=persona,tipo=3,periodo=periodo_h.periodo).last()
                    responsable, facultad, carrera, desde, hasta, estado, search, filtro, url_vars = request.GET.get('responsable', ''),\
                                                                                                      request.GET.get('facultad', ''), \
                                                                                                      request.GET.get('carrera', ''), \
                                                                                                      request.GET.get('desde', ''), \
                                                                                                      request.GET.get('hasta', ''),\
                                                                                                      request.GET.get('estado', ''),\
                                                                                                      request.GET.get('s', ''), \
                                                                                                      Q(status=True, periodo_h=periodo_h), f'&action={action}&id={encrypt(id)}'

                    if es_director:
                        carreras_director=persona.mis_carreras_tercer_nivel().values_list('id', flat=True)
                        filtro = filtro & Q(inscripcion__carrera_id__in=carreras_director, revision_gacademica=1)
                    elif es_director1:
                        filtro = filtro & Q(persona_director=persona, revision_gacademica=1)
                    elif es_asistente:
                        list_carreras=[]
                        for coordinacion in es_asistente.coordinaciones.all():
                            carreras = coordinacion.carrera.all().values_list('id', flat=True)
                            list_carreras+=carreras
                        filtro = filtro & Q(inscripcion__carrera_id__in=list_carreras,revision_director=1)

                    if estado:
                        data['est_g'] = estado = int(estado)
                        url_vars += f"&estado={estado}"
                        if estado == 1:
                            filtro = filtro & (Q(estado=estado) | Q(estado=4))
                        else:
                            filtro = filtro & Q(estado=estado)

                    if desde:
                        data['desde'] = desde
                        url_vars += "&desde={}".format(desde)
                        filtro = filtro & Q(fecha_creacion__gte=desde)
                    if hasta:
                        data['hasta'] = hasta
                        url_vars += "&hasta={}".format(hasta)
                        filtro = filtro & Q(fecha_creacion__lte=hasta)
                    if facultad:
                        data['facultad'] = int(facultad)
                        url_vars += f"&facultad={facultad}"
                        filtro = filtro & Q(inscripcion__coordinacion_id=facultad)
                    if carrera:
                        data['carrera'] = int(carrera)
                        url_vars += f"&carrera={carrera}"
                        filtro = filtro & Q(inscripcion__carrera_id=carrera)
                    if search:
                        data['s'] = search
                        s = search.split(' ')
                        if len(s) == 1:
                            filtro = filtro & (Q(inscripcion__persona__nombres__unaccent__icontains=search) |
                                               Q(inscripcion__persona__cedula__icontains=search) |
                                               Q(inscripcion__persona__apellido1__unaccent__icontains=search) |
                                               Q(inscripcion__persona__apellido2__unaccent__icontains=search))
                        if len(s) >= 2:
                            filtro = filtro & (Q(inscripcion__persona__apellido1__unaccent__icontains=s[0]) & Q(inscripcion__persona__apellido2__unaccent__icontains=s[1]))
                        url_vars += '&s=' + search
                    if responsable and responsable != '0':
                        data['responsable'] = int(responsable)
                        url_vars += f"&responsable={responsable}"
                        filtro = filtro & Q(persona_gacademica_id=responsable)
                    elif responsable_ga and responsable_ga.rol == 0 and not responsable:
                        data['responsable'] = r_ga = responsable_ga.id
                        url_vars += f"&responsable={r_ga}"
                        filtro = filtro & Q(persona_gacademica_id=r_ga)

                    listado = SolicitudEstudianteHomologacion.objects.filter(filtro).order_by('id')
                    paging = MiPaginador(listado, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data["url_vars"] = url_vars
                    data['listado'] = page.object_list

                    data['responsables'] = ResponsableHomologacion.objects.filter(status=True, rol=0)
                    data['estados_g'] = ESTADO_SOLICITUD
                    data['persona'] = persona
                    data['coordinaciones'] = Coordinacion.objects.filter(id__in=[1, 2, 3, 4, 5])
                    data['l_total'] = total= len(listado.values_list('id'))
                    data['l_pendientes'] = len(listado.filter(estado=0).values_list('id'))
                    data['l_corregir'] = len(listado.filter(revision_gacademica=3).values_list('id'))
                    data['l_aprobados'] = len(listado.filter(estado=1).values_list('id'))
                    data['l_rechazados'] = len(listado.filter(estado=2).values_list('id'))
                    data['l_finalizados'] = len(listado.filter(estado=3).values_list('id'))
                    request.session['viewactivo'] = 1
                    if 'exportar_excel' in request.GET:
                        wb = openxl.Workbook()
                        wb["Sheet"].title = "Reporte_solicitudes"
                        ws = wb.active
                        style_title = openxlFont(name='Arial', size=16, bold=True)
                        style_cab = openxlFont(name='Arial', size=10, bold=True)
                        alinear = alin(horizontal="center", vertical="center")
                        response = HttpResponse(content_type="application/ms-excel")
                        response[
                            'Content-Disposition'] = 'attachment; filename=Reporte de solicitudes de homologacion de asignaturas' + '-' + random.randint(
                            1, 10000).__str__() + '.xlsx'
                        ws.column_dimensions['B'].width = 25
                        ws.column_dimensions['C'].width = 15
                        ws.column_dimensions['D'].width = 20
                        ws.column_dimensions['E'].width = 10
                        ws.column_dimensions['F'].width = 25
                        ws.column_dimensions['G'].width = 20
                        ws.column_dimensions['I'].width = 25
                        ws.merge_cells('A1:I1')
                        ws['A1'] = 'REPORTE DE HOMOLOGACIÓN DE ASIGNATURAS'
                        celda1 = ws['A1']
                        celda1.font = style_title
                        celda1.alignment = alinear

                        columns = [u"N°", u"NOMBRES Y APELLIDOS", u"CÉDULA", u"FACULTAD",
                                   u"CARRERA ACTUAL", u"CARRERA HOMOLOGAR", u"N° REVISIONES", u"ESTADO", "RESPONSABLE"
                                   ]
                        row_num = 3
                        for col_num in range(0, len(columns)):
                            celda = ws.cell(row=row_num, column=(col_num + 1), value=columns[col_num])
                            celda.font = style_cab
                        row_num = 4
                        numero=0
                        for list in listado:
                            numero+=1
                            seguimiento = list.seguimiento_revision().exclude(estado_doc=4)
                            carrera = list.carrera_anterior
                            if list.malla_anterior:
                                carrera = list.malla_anterior.carrera
                            ws.cell(row=row_num, column=1, value=numero)
                            ws.cell(row=row_num, column=2, value=str(list.inscripcion.persona))
                            ws.cell(row=row_num, column=3, value=str(list.inscripcion.persona.cedula))
                            ws.cell(row=row_num, column=4, value=str(list.inscripcion.coordinacion))
                            ws.cell(row=row_num, column=5, value=str(list.inscripcion.carrera))
                            ws.cell(row=row_num, column=6, value=str(carrera))
                            ws.cell(row=row_num, column=7, value=len(seguimiento))
                            ws.cell(row=row_num, column=8, value=str(list.get_estado_display()))
                            ws.cell(row=row_num, column=9, value=str(list.persona_gacademica.persona))
                            row_num += 1
                        wb.save(response)
                        return response

                    return render(request, 'adm_homologacion/viewsolicitudes.html', data)
                except Exception as ex:
                    return render({'result': False, 'mensaje': '{}'.format(ex)})

            elif action == 'listcarreras':
                try:
                    idc, idm=int(request.GET.get('id',0)), int(request.GET.get('idm',0))
                    if idc > 0:
                        coordinacion = Coordinacion.objects.get(pk=request.GET['id'])
                        carreras=coordinacion.carreras()
                        if idm > 0:
                           carreras=carreras.filter(modalidad=idm)
                    elif idm > 0:
                        carreras = Carrera.objects.filter(modalidad=idm)
                    lista = []
                    for carrera in carreras:
                        lista.append([carrera.id, f'{carrera}'])
                    return JsonResponse({'result': True, 'lista': lista})
                except Exception as ex:
                    return JsonResponse({"result": False, "mensaje": u"Error al obtener los datos."})

            if action == 'validarrequisitos':
                try:
                    form = ValidarRequisitosForm()
                    data['form'] = form
                    data['filtro'] = SolicitudEstudianteHomologacion.objects.get(pk=request.GET['id'])
                    data['estados'] = ESTADO_VALIDACION_REQUISITOS[1:4]
                    template = get_template("adm_homologacion/modal/formvalidarrequisitos.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            if action == 'validar_director':
                try:
                    data['id'] = id = int(encrypt(request.GET['id']))
                    data['filtro'] = filtro = SolicitudEstudianteHomologacion.objects.get(pk=id)
                    form = ValidarDirectorForm(model_to_dict(filtro))
                    data['form'] = form
                    data['paso'] = 1
                    template = get_template("adm_homologacion/modal/formwizardhomologa.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            if action == 'subirresolucion':
                try:
                    data['id'] = id = int(encrypt(request.GET['id']))
                    data['filtro'] = filtro = SolicitudEstudianteHomologacion.objects.get(pk=id)
                    form = SubirResolucionForm(model_to_dict(filtro))
                    data['form'] = form
                    data['paso'] = 2
                    template = get_template("adm_homologacion/modal/formwizardhomologa.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            if action == 'verproceso':
                try:
                    data['id'] = id = int(encrypt(request.GET['id']))
                    data['filtro'] = filtro = SolicitudEstudianteHomologacion.objects.get(pk=id)
                    data['paso'] = 0
                    template = get_template("adm_homologacion/modal/formwizardhomologa.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            if action == 'verseguimiento':
                try:
                    data['filtro'] = SolicitudEstudianteHomologacion.objects.get(pk=int(encrypt(request.GET['id'])))
                    template = get_template('adm_homologacion/modal/formseguimiento.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, "mensaje": ex})

            #Reporteria
            elif action == 'descargarexpediente':
                try:
                    id = int(encrypt(request.GET['id']))
                    documentos = DocumentosSolicitudHomologacion.objects.filter(status=True, solicitud_id=id).distinct()
                    solicitud = documentos.first().solicitud
                    name_user = solicitud.inscripcion.persona.usuario.username
                    name_zip=f'exp_homologacion_{name_user}_{random.randint(1, 10000).__str__()}.zip'
                    directory = os.path.join(SITE_STORAGE,'media', 'expedientes_homologacion')
                    try:
                        os.stat(directory)
                    except:
                        os.mkdir(directory)

                    url = os.path.join(SITE_STORAGE, 'media', 'expedientes_homologacion',name_zip)
                    fantasy_zip = zipfile.ZipFile(url, 'w')
                    cont=0
                    for inf in documentos:
                        if inf.archivo:
                            cont+=1
                            name_doc = remover_caracteres_especiales_unicode(inf.name_documento().__str__().lower().replace(' ', '_')).lower().replace(' ', '_')
                            fantasy_zip.write(inf.archivo.path,f'{cont}_{name_user}_{name_doc}.pdf')
                    if solicitud.archivoresoluciondirector:
                        fantasy_zip.write(solicitud.archivoresoluciondirector.path,f'{cont + 1}_{name_user}_resolucion_director_carrera.pdf')
                    if solicitud.archivoresoluciondirectivo:
                        fantasy_zip.write(solicitud.archivoresoluciondirectivo.path,f'{cont + 2}_{name_user}_resolucion_directivo.pdf')
                    fantasy_zip.close()
                    response = HttpResponse(open(url, 'rb'), content_type='application/zip')
                    response['Content-Disposition'] = f'attachment; filename={name_zip}'
                    return response
                except Exception as ex:
                    print('Error en linea {}'.format(sys.exc_info()[-1].tb_lineno))
                    messages.error(request, f'{ex}')
                    return redirect(f'{request.path}?action=solicitudes&id={encrypt(solicitud.periodo_h.id)}')

            elif action == 'descargarrevision':
                try:
                    wb = openxl.Workbook()
                    wb["Sheet"].title = "Historial de revisión"
                    ws = wb.active
                    style_title = openxlFont(name='Arial', size=16, bold=True)
                    style_cab = openxlFont(name='Arial', size=10, bold=True)
                    alinear = alin(horizontal="center", vertical="center")
                    id_s=int(encrypt(request.GET['id']))
                    solicitud=SolicitudEstudianteHomologacion.objects.get(id=id_s)
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = f'attachment; filename=Proceso de revisión - {id_s}' + '-' + random.randint(1, 10000).__str__() + '.xlsx'
                    ws.column_dimensions['B'].width = 30
                    ws.column_dimensions['C'].width = 15
                    ws.column_dimensions['D'].width = 15
                    ws.column_dimensions['E'].width = 25
                    ws.column_dimensions['F'].width = 35
                    ws.merge_cells('A1:F1')
                    ws['A1'] = 'HISTORIAL DE REVISIÓN DE HOMOLOGACIÓN DE ASIGNATURAS'
                    celda1 = ws['A1']
                    celda1.font = style_title
                    celda1.alignment = alinear
                    ws['A2'] = 'NOMBRE'
                    ws['A2'].font = style_cab
                    ws['B2'] = str(solicitud.inscripcion.persona.nombre_completo_minus())
                    ws['C2'] = 'CÉDULA'
                    ws['C2'].font = style_cab
                    ws['D2'] = str(solicitud.inscripcion.persona.cedula)
                    ws['E2'] = 'CARRERA ACTUAL'
                    ws['E2'].font = style_cab
                    ws['F2'] = str(solicitud.inscripcion.carrera)
                    columns = [u"N°", u"DOCUMENTO", u"FECHA REVISIÓN", u"ESTADO", u"QUIEN REVISO", "OBSERVACIÓN"
                               ]
                    row_num = 3
                    for col_num in range(0, len(columns)):
                        celda = ws.cell(row=row_num, column=(col_num + 1), value=columns[col_num])
                        celda.font = style_cab
                    listado = SeguimientoRevision.objects.filter(solictud_id=id_s, documento__isnull=False, revisor__isnull=False)
                    mensaje = 'No registra'
                    row_num = 4
                    numero = 0
                    for list in listado:
                        numero+=1
                        ws.cell(row=row_num, column=1, value=numero)
                        ws.cell(row=row_num, column=2, value=str(list.documento.name_documento()))
                        ws.cell(row=row_num, column=3, value=str(list.fecha_creacion.strftime('%d-%m-%Y')))
                        ws.cell(row=row_num, column=4, value=str(list.get_estado_doc_display()))
                        ws.cell(row=row_num, column=5, value=str(list.revisor.nombre_completo_minus()))
                        ws.cell(row=row_num, column=6, value=str(list.observacion) if list.observacion else mensaje)
                        row_num += 1
                    wb.save(response)
                    return response
                except Exception as ex:
                    print('Error en linea {}'.format(sys.exc_info()[-1].tb_lineno))
                    messages.error(request, f'{ex}')
                    return redirect(f'{request.path}?action=solicitudes&id={encrypt(solicitud.periodo_h.id)}')

            return HttpResponseRedirect(request.path)
        else:
            try:
                data['title'] = u'Periodos Homologación'
                search, filtro, url_vars = request.GET.get('s', ''), Q(status=True), ''

                if search:
                    filtro = filtro & (Q(periodo__nombre__unaccent__icontains=search))
                    url_vars += '&s=' + search
                    data['s'] = search

                listado = PeriodoHomologacion.objects.filter(filtro)
                paging = MiPaginador(listado, 10)
                p = 1
                try:
                    paginasesion = 1
                    if 'paginador' in request.session:
                        paginasesion = int(request.session['paginador'])
                    if 'page' in request.GET:
                        p = int(request.GET['page'])
                    else:
                        p = paginasesion
                    try:
                        page = paging.page(p)
                    except:
                        p = 1
                    page = paging.page(p)
                except:
                    page = paging.page(p)
                request.session['paginador'] = p
                data['paging'] = paging
                data['rangospaging'] = paging.rangos_paginado(p)
                data['page'] = page
                data["url_vars"] = url_vars
                data['listado'] = page.object_list
                data['listcount'] = total= len(listado)
                request.session['viewactivo'] = 1
                return render(request, 'adm_homologacion/viewperiodohomologacion.html', data)
            except Exception as ex:
                print('Error en linea {}'.format(sys.exc_info()[-1].tb_lineno))
                messages.error(request, f'{ex}')
                return redirect(f'{request.path}')
