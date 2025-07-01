# -*- coding: UTF-8 -*-
import random
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Q
from django.http import HttpResponseRedirect, JsonResponse, HttpResponse
from django.shortcuts import render
from decorators import secure_module
from sagest.forms import ProveedorForm
from sagest.models import Proveedor
from settings import EMAIL_DOMAIN
from sga.commonviews import adduserdata
from sga.funciones import MiPaginador, log, puede_realizar_accion

from openpyxl import workbook as openxl
from openpyxl.styles import Font as openxlFont
from openpyxl.styles.alignment import Alignment as alin


@login_required(redirect_field_name='ret', login_url='/loginsagest')
@secure_module
@transaction.atomic()
def view(request):
    data = {}
    adduserdata(request, data)
    if request.method == 'POST':
        action = request.POST['action']

        if action == 'add':
            try:
                f = ProveedorForm(request.POST)
                if f.is_valid():
                    if Proveedor.objects.filter(identificacion=f.cleaned_data['identificacion'].strip()).exists():
                        return JsonResponse({"result": "bad", "mensaje": u"Ya existe un proveedor registrado con ese número de identificacion."})
                    proveedor = Proveedor(identificacion=f.cleaned_data['identificacion'],
                                          nombre=f.cleaned_data['nombre'],
                                          alias=f.cleaned_data['alias'],
                                          pais=f.cleaned_data['pais'],
                                          direccion=f.cleaned_data['direccion'],
                                          telefono=f.cleaned_data['telefono'],
                                          celular=f.cleaned_data['celular'],
                                          email=f.cleaned_data['email'],
                                          fax=f.cleaned_data['fax'])
                    proveedor.save(request)
                    log(u'Adiciono nuevo proveedor: %s' % proveedor, request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                     raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        if action == 'edit':
            try:
                proveedor = Proveedor.objects.get(pk=request.POST['id'])
                f = ProveedorForm(request.POST)
                if f.is_valid():
                    proveedor.alias = f.cleaned_data['alias']
                    proveedor.direccion = f.cleaned_data['direccion']
                    proveedor.pais = f.cleaned_data['pais']
                    proveedor.telefono = f.cleaned_data['telefono']
                    proveedor.celular = f.cleaned_data['celular']
                    proveedor.email = f.cleaned_data['email']
                    proveedor.fax = f.cleaned_data['fax']
                    proveedor.save(request)
                    log(u'Modificó proveedor: %s' % proveedor, request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                     raise NameError('Error')
            except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        if action == 'delete':
            try:
                proveedor = Proveedor.objects.get(pk=request.POST['id'])
                if proveedor.en_uso():
                    return JsonResponse({"result": "bad", "mensaje": u"El proveedor se encuentra en uso, no es posible eliminar."})
                log(u'Eliminó proveedor: %s' % proveedor, request, "del")
                proveedor.delete()
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        if action == 'generar_excel':
            try:
                proveedores = Proveedor.objects.filter(status=True)
                wb = openxl.Workbook()
                wb["Sheet"].title = "Resumen_proveedores"
                ws = wb.active
                style_title = openxlFont(name='Arial', size=16, bold=True)
                style_cab = openxlFont(name='Arial', size=10, bold=True)
                alinear = alin(horizontal="center", vertical="center")
                response = HttpResponse(content_type="application/ms-excel")
                response[
                    'Content-Disposition'] = 'attachment; filename=Resumen de Proveedores' + '-' + random.randint(
                    1, 10000).__str__() + '.xlsx'
                ws.column_dimensions['B'].width = 15
                ws.column_dimensions['C'].width = 15
                ws.column_dimensions['D'].width = 40
                ws.column_dimensions['E'].width = 30
                ws.column_dimensions['F'].width = 30
                ws.column_dimensions['G'].width = 30
                ws.column_dimensions['H'].width = 35
                ws.column_dimensions['I'].width = 25
                ws.column_dimensions['J'].width = 25

                ws.merge_cells('A1:J2')
                ws['A1'] = 'UNIVERSIDAD ESTATAL DE MILAGRO\n DEPARTAMENTO FINANCIERO'
                celda1 = ws['A1']
                celda1.font = style_title
                celda1.alignment = alinear

                ws.merge_cells('A3:B3')
                ws['A3'] = 'Resumen Proveedores'
                celda1 = ws['A3']
                celda1.font = style_cab

                columns = [u"Identificación", u"Nombre", u"Alias", u"Pais / Dirección", u"Telefóno",
                           u"Celular", u"Email"]
                row_num = 5
                for col_num in range(0, len(columns)):
                    celda = ws.cell(row=row_num, column=(col_num + 1), value=columns[col_num])
                    celda.font = style_cab
                row_num = 6
                numero = 1
                for proveedor in proveedores:
                    ws.cell(row=row_num, column=1, value=numero)
                    ws.cell(row=row_num, column=2, value=str(proveedor.identificacion))
                    ws.cell(row=row_num, column=3, value=str(proveedor.nombre))
                    ws.cell(row=row_num, column=4, value=str(proveedor.alias))
                    ws.cell(row=row_num, column=5,
                            value=str(proveedor.pais) if proveedor.pais else ' ' + str(proveedor.direccion))
                    ws.cell(row=row_num, column=6,
                            value=str(proveedor.celular) if proveedor.celular else '')
                    ws.cell(row=row_num, column=7, value=str(proveedor.telefono) if proveedor.telefono else '')
                    ws.cell(row=row_num, column=8, value=str(proveedor.email))
                    row_num += 1
                    numero += 1
                wb.save(response)
                return response


            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        return JsonResponse({"result": "bad", "mensaje": u"Solicitud Incorrecta."})
    else:
        if 'action' in request.GET:
            action = request.GET['action']

            if action == 'add':
                try:
                    puede_realizar_accion(request, 'sagest.puede_modificar_proveedor')
                    data['title'] = u'Adicionar Proveedor'
                    data['form'] = ProveedorForm()
                    return render(request, "adm_proveedores/add.html", data)
                except Exception as ex:
                    pass

            if action == 'edit':
                try:
                    puede_realizar_accion(request, 'sagest.puede_modificar_proveedor')
                    data['title'] = u'Editar Proveedor'
                    data['proveedor'] = proveedor = Proveedor.objects.get(pk=request.GET['id'])
                    form = ProveedorForm(initial={'identificacion': proveedor.identificacion,
                                                  'nombre': proveedor.nombre,
                                                  'alias': proveedor.alias,
                                                  'direccion': proveedor.direccion,
                                                  'pais': proveedor.pais,
                                                  'telefono': proveedor.telefono,
                                                  'celular': proveedor.celular,
                                                  'email': proveedor.email,
                                                  'fax': proveedor.fax})
                    form.editar()
                    data['form'] = form
                    return render(request, "adm_proveedores/edit.html", data)
                except Exception as ex:
                    pass

            if action == 'delete':
                try:
                    puede_realizar_accion(request, 'sagest.puede_modificar_proveedor')
                    data['title'] = u'Borrar Proveedor'
                    data['proveedor'] = Proveedor.objects.get(pk=request.GET['id'])
                    return render(request, "adm_proveedores/delete.html", data)
                except Exception as ex:
                    pass
            return HttpResponseRedirect(request.path)

        else:
            data['title'] = u'Gestión de proveedores'
            search = None
            ids = None
            if 's' in request.GET:
                search = request.GET['s']
            if search:
                proveedores = Proveedor.objects.filter(Q(nombre__icontains=search) |
                                                       Q(alias__icontains=search) |
                                                       Q(identificacion__icontains=search))
            elif 'id' in request.GET:
                ids = request.GET['id']
                proveedores = Proveedor.objects.filter(id=ids)
            else:
                proveedores = Proveedor.objects.all()
            paging = MiPaginador(proveedores, 20)
            p = 1
            try:
                paginasesion = 1
                if 'paginador' in request.session:
                    paginasesion = int(request.session['paginador'])
                if 'page' in request.GET:
                    p = int(request.GET['page'])
                else:
                    p = paginasesion
                try:
                    page = paging.page(p)
                except:
                    p = 1
                page = paging.page(p)
            except:
                page = paging.page(p)
            request.session['paginador'] = p
            data['paging'] = paging
            data['rangospaging'] = paging.rangos_paginado(p)
            data['page'] = page
            data['search'] = search if search else ""
            data['ids'] = ids if ids else ""
            data['proveedores'] = page.object_list
            data['email_domain'] = EMAIL_DOMAIN

            if 'generar_excel' in request.GET:

                proveedores = Proveedor.objects.filter(status=True)
                wb = openxl.Workbook()
                wb["Sheet"].title = "Resumen_proveedores"
                ws = wb.active
                style_title = openxlFont(name='Arial', size=16, bold=True)
                style_cab = openxlFont(name='Arial', size=10, bold=True)
                alinear = alin(horizontal="center", vertical="center")
                response = HttpResponse(content_type="application/ms-excel")
                response[
                    'Content-Disposition'] = 'attachment; filename=Resumen de Proveedores' + '-' + random.randint(
                    1, 10000).__str__() + '.xlsx'
                ws.column_dimensions['B'].width = 15
                ws.column_dimensions['C'].width = 15
                ws.column_dimensions['D'].width = 40
                ws.column_dimensions['E'].width = 30
                ws.column_dimensions['F'].width = 30
                ws.column_dimensions['G'].width = 30
                ws.column_dimensions['H'].width = 35
                ws.column_dimensions['I'].width = 25
                ws.column_dimensions['J'].width = 25

                ws.merge_cells('A1:J2')
                ws['A1'] = 'UNIVERSIDAD ESTATAL DE MILAGRO\n DEPARTAMENTO FINANCIERO'
                celda1 = ws['A1']
                celda1.font = style_title
                celda1.alignment = alinear

                ws.merge_cells('A3:B3')
                ws['A3'] = 'Resumen Proveedores'
                celda1 = ws['A3']
                celda1.font = style_cab

                columns = [u"Identificación", u"Nombre", u"Alias", u"Pais / Dirección", u"Telefóno",
                           u"Celular", u"Email"]
                row_num = 5
                for col_num in range(0, len(columns)):
                    celda = ws.cell(row=row_num, column=(col_num + 1), value=columns[col_num])
                    celda.font = style_cab
                row_num = 6
                numero = 1
                for proveedor in proveedores:
                    ws.cell(row=row_num, column=1, value=str(proveedor.identificacion))
                    ws.cell(row=row_num, column=2, value=str(proveedor.nombre))
                    ws.cell(row=row_num, column=3, value=str(proveedor.alias))
                    ws.cell(row=row_num, column=4,
                            value=str(proveedor.pais) if proveedor.pais else ' ' + str(proveedor.direccion))
                    ws.cell(row=row_num, column=5,
                            value=str(proveedor.celular) if proveedor.celular else '')
                    ws.cell(row=row_num, column=6, value=str(proveedor.telefono) if proveedor.telefono else '')
                    ws.cell(row=row_num, column=7, value=str(proveedor.email))
                    row_num += 1
                    numero += 1
                wb.save(response)
                return response

            return render(request, "adm_proveedores/view.html", data)
