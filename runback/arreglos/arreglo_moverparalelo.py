#!/usr/bin/env python
# -*- coding: utf-8 -*-
import json
import os
import sys

import openpyxl
# import urllib2

# Full path and name to your csv file
import unicodedata
# from django.db.backends.oracle.base import to_unicode
# from apt.package import Record

import xlrd
# from __builtin__ import file
# from IPython.lib.editorhooks import mate
import xlwt
from django.http import HttpResponse
# from numpy.core.records import record
# from numpy.matrixlib.defmatrix import matrix
from setuptools.windows_support import hide_file
from urllib3 import request
from docx import Document
from xlwt import easyxf


# SITE_ROOT = os.path.dirname(os.path.realpath(__file__))
YOUR_PATH = os.path.dirname(os.path.realpath(__file__))
# print(f"YOUR_PATH: {YOUR_PATH}")
SITE_ROOT = os.path.dirname(os.path.dirname(YOUR_PATH))
SITE_ROOT = os.path.join(SITE_ROOT, '')
# print(f"SITE_ROOT: {SITE_ROOT}")
your_djangoproject_home = os.path.split(SITE_ROOT)[0]
# print(f"your_djangoproject_home: {your_djangoproject_home}")
sys.path.append(your_djangoproject_home)

from django.core.wsgi import get_wsgi_application
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "settings")
application = get_wsgi_application()
import xlrd
from time import sleep
from django.template import Context
from sga.models import *
from sagest.models import *
from datetime import date
from settings import PROFESORES_GROUP_ID
from sga.funciones import calculate_username, generar_usuario



# try:
#     workbook = xlrd.open_workbook('mover.xlsx')
#     sheet = workbook.sheet_by_index(0)
#     linea = 1
#     for rowx in range(sheet.nrows):
#         if linea > 1:
#             cols = sheet.row_values(rowx)
#             matricula = Matricula.objects.get(pk=int(cols[0]))
#             carrera = matricula.inscripcion.carrera
#             for matasig in matricula.materiaasignada_set.filter(materia__paralelo=cols[2],status=True):
#                 # materia2 = Materia.cronograma_s
#                 # print(matasig.materia.asignatura.id)
#                 # materia2 = Materia.objects.get(nivel__periodo_id=85, asignatura=matasig.materia.asignatura, asignaturamalla__malla__carrera=carrera,paralelo = cols[3], status=True)
#                 matriculas = matasig.matriculas
#                 carreram= matasig.materia.asignaturamalla.malla.carrera
#                 materia2 = Materia.objects.get(nivel__periodo_id=110, asignatura=matasig.materia.asignatura, asignaturamalla__malla__carrera=carreram,paralelo = cols[3], status=True)
#
#                 #matriculas = matricula.inscripcion.historicorecordacademico_set.values('id').filter(asignatura=materia2.asignatura, fecha__lt=materia2.nivel.fin).count() + 1
#                 if not MateriaAsignada.objects.filter(matricula=matricula,materia=materia2):
#                     asistencias = matasig.asistencialeccion_set.all()
#                     asistencias.delete()
#                     evaluaciones = matasig.evaluacion()
#                     evaluaciones.delete()
#                     planifi = matasig.materiaasignadaplanificacion_set.all()
#                     planifi.delete()
#                     materiaasignada = MateriaAsignada(matricula=matricula,
#                                                       materia=materia2,
#                                                       notafinal=0,
#                                                       asistenciafinal=0,
#                                                       cerrado=False,
#                                                       matriculas=matriculas,
#                                                       # matriculas=1,
#                                                       observaciones='',
#                                                       estado_id=NOTA_ESTADO_EN_CURSO)
#                     materiaasignada.save()
#                     materiaasignada.asistencias()
#                     materiaasignada.evaluacion()
#                     materiaasignada.mis_planificaciones()
#                     materiaasignada.save()
#                     matasig.delete()
#
#                     print(matasig)
#                     print(materia2)
#                     print(linea)
#                     print('-------------')
#
#         linea += 1
#         # print(linea)
# except Exception as ex:
#     print('error : %s' % ex)
#


# materias = Materia.objects.filter(nivel__modalidad_id=3 , nivel__periodo_id=82)
# for lismate in materias:
#     json_content = ''
#     template = ''
#     data = {}
#     numtemas = 0
#     totalcon = int(DetalleSilaboSemanalTema.objects.values('temaunidadresultadoprogramaanalitico_id').filter(temaunidadresultadoprogramaanalitico__status=True, silabosemanal__status=True, silabosemanal__silabo__status=True, silabosemanal__silabo__materia=lismate).distinct().count() / 2)
#     data['numtemas'] = numtemas = list(range(1,totalcon + 1))
#     data['nommateria'] = lismate
#     template = get_template("contenidosonline/tableejemplo.html")
#     json_content = template.render(data)
#     lismate.temas_embebidos = json_content.strip()
#     lismate.save()
#     print(lismate)
# silabocab = Silabo.objects.get(pk=5690, status=True)
# SilaboSemanal.objects.filter(silabo=silabocab).update(status=False)

# de = DetalleDistributivo.objects.get(pk=55558).delete()

#
# silabocab = Silabo.objects.get(pk=5690, status=True)
# for p in PlanificacionClaseSilabo.objects.filter(
#         tipoplanificacion__planificacionclasesilabo_materia__materia=silabocab.materia, status=True).exclude(
#         semana=0).order_by('orden'):
#     if silabocab.silabosemanal_set.filter(fechainiciosemana__gte=p.fechainicio, fechafinciosemana__lte=p.fechafin).exists():
#         lissemana = silabocab.silabosemanal_set.filter(fechainiciosemana__gte=p.fechainicio, fechafinciosemana__lte=p.fechafin)[0]
#         idcodigo = lissemana.id
#         semana = lissemana
#         lissemana.status=True
#         lissemana.save()

#
# materiaasignada = MateriaAsignada.objects.filter(materia__nivel__periodo__id=85)
# for mate in materiaasignada:
#     if mate.asistencialeccion_set.filter(leccion__fecha__gte=mate.fechaasignacion).exists():
#         if mate.matricula.inscripcion.carrera.coordinacion_set.filter(status=True)[0].id == 7:
#             # solo postgrado
#             total = mate.asistencia_plan()
#         else:
#             total = mate.asistencialeccion_set.values("id").filter(leccion__fecha__gte=mate.fechaasignacion,
#                                                                    leccion__fecha__lte=mate.materia.fechafinasistencias).distinct().count()
#         if total:
#             real = mate.asistencialeccion_set.values("id").filter(leccion__fecha__gte=mate.fechaasignacion,
#                                                                   leccion__fecha__lte=mate.materia.fechafinasistencias,
#                                                                   asistio=True).exclude(
#                 justificacionausenciaasistencialeccion__isnull=False).distinct().count()
#             justificada = null_to_numeric(JustificacionAusenciaAsistenciaLeccion.objects.filter(
#                 asistencialeccion__leccion__fecha__gte=mate.fechaasignacion,
#                 asistencialeccion__leccion__fecha__lte=mate.materia.fechafinasistencias,
#                 asistencialeccion__materiaasignada=mate).distinct().aggregate(cantidad=Sum('porcientojustificado'))[
#                                               'cantidad'])
#             mate.asistenciafinal = round(((real + justificada) * 100) / float(total), 0)
#             mate.save()
#     print(mate)
#
# listado = DetalleDistributivo.objects.filter(distributivo__periodo=85,criteriodocenciaperiodo_id__in=[237,238,240,239], actividaddetalledistributivo__isnull=True, criteriodocenciaperiodo__isnull=False, status=True)
# contador = 0
# for lis in listado:
#     contador = contador + 1
#     cri = lis.criteriodocenciaperiodo.criterio.nombre
#     acti = ActividadDetalleDistributivo(criterio=lis,
#                                         nombre=cri,
#                                         desde='2019-05-31',
#                                         hasta='2019-09-29',
#                                         horas=lis.horas)
#     acti.save()
#     print(contador)


# import datetime
# # datetime.datetime.today()
# fepunemi = str(datetime.datetime.now().date())
# # fechaepunemi = datetime.strptime(fepunemi, "%Y-%m-%d").date()
# # datetime.datetime(2012, 3, 23, 23, 24, 55, 173504)
# datetime.datetime.today().weekday()
# # print(datetime.datetime.today().weekday())
# # print(datetime.datetime.now().date())
# print(datetime.datetime.strptime(fepunemi, "%Y-%m-%d").date().weekday()+1)

# acciones = AccionDocumentoDetalle.objects.filter(acciondocumento__status=True, mostrar=True, acciondocumento__indicadorpoa__objetivooperativo__objetivotactico__objetivoestrategico__periodopoa_id=8,  status=True)
# a=0
# for acc in acciones:
#     if acc.acciondocumentodetallerecord_set.exists():
#         rec = acc.acciondocumentodetallerecord_set.order_by("-id")[0]
#         if rec.procesado:
#             a = a+1
#             rec.acciondocumentodetalle.estado_rubrica = rec.rubrica_aprobacion
#             rec.acciondocumentodetalle.estado_rubrica = rec.rubrica_aprobacion
#             rec.acciondocumentodetalle.mostrar = True
#             rec.acciondocumentodetalle.save()
# print(a)
# def lista_detalles_revisiondos(self):
#     lista = []
#     cantidad_meses = 0
#     if self.indicadorpoa.objetivooperativo.objetivotactico.objetivoestrategico.periodopoa.anio < datetime.now().year:
#         mes_revision = 12
#     else:
#         mes_revision = datetime.now().month
#     ingresar = self.indicadorpoa.objetivooperativo.objetivotactico.objetivoestrategico.periodopoa.ingresar
#     for m in range(1, 13):
#         if self.acciondocumentodetalle_set.filter(inicio__month=m, status=True).exists():
#             det = self.acciondocumentodetalle_set.filter(inicio__month=m, status=True)[0]
#             ok = evidencia = revisado = por_revisar = record = bloqueo = 0
#             action = ""
#             if mes_revision > m or mes_revision == 12:
#                 detrecorok = det.acciondocumentodetallerecord_set.order_by("-id")
#                 if detrecorok.exists():
#                     evidencia = 1
#                     record = detrecorok[0].id
#                     if detrecorok[0].usuario_revisa is not None and detrecorok[0].usuario_aprobacion is None:
#                         revisado = 1
#                     elif detrecorok[0].usuario_revisa is None and detrecorok[0].usuario_aprobacion is None:
#                         por_revisar = 1
#                     elif detrecorok[0].usuario_revisa is not None and detrecorok[0].usuario_aprobacion is None:
#                         bloqueo = 0
#                     elif detrecorok[0].usuario_revisa is not None and detrecorok[0].usuario_aprobacion is not None:
#                         bloqueo = 1
#                     if not ingresar:
#                         bloqueo = 1
#                     if detrecorok[0].usuario_envia is None:
#                         action = "sin_evidenciados"
#                     else:
#                         action = "con_evidenciados"
#                 else:
#                     if not ingresar:
#                         bloqueo = 1
#                     tipo = "addsin"
#                     action = "sin_evidenciados"
#                 ok = 1
#             if not ingresar and evidencia == 0:
#                 ok = 0
#             mes_inicio = det.inicio.month
#             mes_fin = det.fin.month
#             cantidad_meses = mes_fin - mes_inicio
#             # lista.append((det.id, det.inicio, det.fin, cantidad_meses + 1, ok, revisado, por_revisar, evidencia, record, det.get_estado_accion_display(), action, det.estado_accion, bloqueo))
#             lista.append((det.id, det.inicio, det.fin, cantidad_meses + 1, ok, revisado, por_revisar, evidencia, record,
#                           det.estado_rubrica.nombre, action, det.estado_rubrica.id, bloqueo))
#         else:
#             if not lista or cantidad_meses == 0:
#                 lista.append((0, '-', '-', 1, 0, 0, 0, 0, 0, "", "", 0, 0))
#             else:
#                 cantidad_meses -= 1
#     return lista


# try:
#     workbook = xlrd.open_workbook('criterios.xlsx')
#     sheet = workbook.sheet_by_index(0)
#     linea = 1
#     for rowx in range(sheet.nrows):
#         if linea > 1:
#             cols = sheet.row_values(rowx)
#             actividad = ActividadPrincipal(nombre=cols[1],
#                                            tipocriterioactividadprincipal=int(cols[0]),
#                                            vigente=True)
#             actividad.save()
#
#         linea += 1
#         print(linea)
# except Exception as ex:
#     print('error: %s' % ex)

# listadistributivo = ProfesorDistributivoHoras.objects.filter(periodo_id=90).exclude(profesor__persona__apellido1__icontains='APELLIDOVIRTUAL').exclude(profesor__persona__nombres__icontains='POR DEFINIR').exclude(profesor__persona__apellido1__icontains='POR DEFINIR').exclude(profesor__persona__apellido2__icontains='POR DEFINIR')
# cuenta = 0
# for lista in listadistributivo:
#     cuenta += 1
#     if not ProfesorDistributivoHoras.objects.filter(periodo_id=97,profesor=lista.profesor):
#         distri = ProfesorDistributivoHoras(periodo_id=97,
#                                            profesor=lista.profesor,
#                                            dedicacion=lista.dedicacion,
#                                            categoria=lista.categoria,
#                                            cargo=lista.cargo,
#                                            nivelcategoria=lista.nivelcategoria,
#                                            nivelescalafon=lista.nivelescalafon,
#                                            coordinacion=lista.coordinacion,
#                                            carrera=lista.carrera,
#                                            activo=lista.activo
#                                            )
#         distri.save()
#     print(cuenta)

# listacriterio = CriterioDocenciaPeriodo.objects.filter(periodo_id=96, status=True)
# cuenta = 0
# for crite in listacriterio:
#     cuenta += 1
#     if not CriterioDocenciaPeriodo.objects.filter(periodo_id=97, criterio=crite.criterio, status=True):
#         distri = CriterioDocenciaPeriodo(periodo_id=97,
#                                          criterio=crite.criterio,
#                                          minimo=crite.minimo,
#                                          maximo=crite.maximo,
#                                          articulo=crite.articulo,
#                                          actividad=crite.actividad
#                                          )
#         distri.save()
#     else:
#         docen = CriterioDocenciaPeriodo.objects.get(periodo_id=97, criterio=crite.criterio, status=True)
#         docen.minimo=crite.minimo
#         docen.maximo=crite.maximo
#         docen.save()
#
#     print(cuenta)
#
# listacriterio = CriterioInvestigacionPeriodo.objects.filter(periodo_id=96, status=True)
# cuenta = 0
# for crite in listacriterio:
#     cuenta += 1
#     if not CriterioInvestigacionPeriodo.objects.filter(periodo_id=97, criterio=crite.criterio, status=True):
#         distri = CriterioInvestigacionPeriodo(periodo_id=97,
#                                          criterio=crite.criterio,
#                                          minimo=crite.minimo,
#                                          maximo=crite.maximo,
#                                          articulo=crite.articulo,
#                                          productoinvestigacion=crite.productoinvestigacion,
#                                          actividad=crite.actividad
#                                          )
#         distri.save()
#         print(cuenta)
#
# listacriterio = CriterioGestionPeriodo.objects.filter(periodo_id=96, status=True)
# cuenta = 0
# for crite in listacriterio:
#     cuenta += 1
#     if not CriterioGestionPeriodo.objects.filter(periodo_id=97, criterio=crite.criterio, status=True):
#         distri = CriterioGestionPeriodo(periodo_id=97,
#                                          criterio=crite.criterio,
#                                          minimo=crite.minimo,
#                                          maximo=crite.maximo,
#                                          articulo=crite.articulo,
#                                          es_admision=crite.es_admision,
#                                          actividad=crite.actividad
#                                          )
#         distri.save()
#         print(cuenta)

# listadovaloracion = DetalleMatrizValoracionPoa.objects.filter(matrizvaloracion__evaluacionperiodo_id=1, status=True)
# for lis in listadovaloracion:
#     if not DetalleMatrizEvaluacionPoa.objects.filter(matrizvaloracion=lis.matrizvaloracion, actividad=lis.actividad, status=True):
#         detalleeval = DetalleMatrizEvaluacionPoa(matrizvaloracion=lis.matrizvaloracion,
#                                                  actividad=lis.actividad,
#                                                  estado_rubrica=lis.estado_rubrica,
#                                                  cumplimiento=lis.cumplimiento,
#                                                  descripcion=lis.descripcion,
#                                                  semanaplanificada=lis.semanaplanificada,
#                                                  semanaejecutada=lis.semanaejecutada,
#                                                  cumplimientosemana=lis.cumplimientosemana,
#                                                  presupuestoreformado=lis.presupuestoreformado,
#                                                  presupuestoutilizado=lis.presupuestoutilizado,
#                                                  indicadoreficacia=lis.indicadoreficacia,
#                                                  indicadoreficienciatiempo=lis.indicadoreficienciatiempo,
#                                                  indicadoreficienciapresupuesto=lis.indicadoreficienciapresupuesto,
#                                                  indicadoreficiencia=lis.indicadoreficiencia,
#                                                  indicadordesempeno=lis.indicadordesempeno,
#                                                  metaejecutada=lis.metaejecutada,
#                                                  cumplimientometa=lis.cumplimientometa,
#                                                  observacion=lis.observacion,
#                                                  recomendacion=lis.recomendacion)
#         detalleeval.save()
#     print(lis.id)

# listado = MatrizValoracionPoa.objects.filter(evaluacionperiodo_id=2, status=True)
#
# for lis in listado:
#     if lis.detallematrizvaloracionpoa_set.filter(status=True):
#         lista1 = AccionDocumentoDetalle.objects.filter(acciondocumento__indicadorpoa__objetivooperativo__objetivotactico__objetivoestrategico__departamento_id=lis.departamento.id,
#                                                        acciondocumento__indicadorpoa__objetivooperativo__objetivotactico__objetivoestrategico__departamento__status=True,
#                                                        acciondocumento__indicadorpoa__objetivooperativo__objetivotactico__objetivoestrategico__periodopoa_id=lis.evaluacionperiodo.periodopoa.id,
#                                                        acciondocumento__indicadorpoa__objetivooperativo__objetivotactico__objetivoestrategico__periodopoa__status=True,
#                                                        acciondocumento__status=True,status=True, acciondocumentodetallerecord__procesado=False, inicio__year=lis.evaluacionperiodo.fechainicio.year)
#         lista3 = AccionDocumentoDetalle.objects.filter(acciondocumento__indicadorpoa__objetivooperativo__objetivotactico__objetivoestrategico__departamento_id=lis.departamento.id,
#                                                        acciondocumento__indicadorpoa__objetivooperativo__objetivotactico__objetivoestrategico__departamento__status=True,
#                                                        acciondocumento__indicadorpoa__objetivooperativo__objetivotactico__objetivoestrategico__periodopoa_id=lis.evaluacionperiodo.periodopoa.id,
#                                                        acciondocumento__indicadorpoa__objetivooperativo__objetivotactico__objetivoestrategico__periodopoa__status=True,
#                                                        acciondocumento__status=True, status=True, acciondocumentodetallerecord__procesado=True, acciondocumentodetallerecord__rubrica_aprobacion_id=3, inicio__year=lis.evaluacionperiodo.fechainicio.year).exclude(acciondocumento__indicadorpoa__id__in=lista1.values_list('acciondocumento__indicadorpoa__id'))
#
#         for tres in lista3:
#             # if not lis.detallematrizvaloracionpoa_set.filter(actividad_id=tres.acciondocumento.indicadorpoa.id, status=True):
#             #     detallematriz = DetalleMatrizValoracionPoa(matrizvaloracion=lis,
#             #                                                actividad_id=tres.acciondocumento.indicadorpoa.id,
#             #                                                estado_rubrica_id=3)
#             #     detallematriz.save()
#             print(str(tres.acciondocumento.indicadorpoa.id) + ' - ' + str(lis.departamento))
#             idactividadvaloracion = DetalleMatrizValoracionPoa.objects.get(matrizvaloracion=lis, actividad=tres.acciondocumento.indicadorpoa , status=True)
#             if lis.detallematrizevaluacionpoa_set.filter(actividad=tres.acciondocumento.indicadorpoa, status=True):
#                 detalle = DetalleMatrizEvaluacionPoa.objects.get(matrizvaloracion=lis, actividad=tres.acciondocumento.indicadorpoa,  status=True)
#                 detalle.estado_rubrica = idactividadvaloracion.estado_rubrica
#                 detalle.cumplimiento = idactividadvaloracion.cumplimiento
#                 detalle.save()
#                 print(str(detalle.estado_rubrica) + ' ' + str(idactividadvaloracion.estado_rubrica))

# from django.http import HttpResponse
# from xlwt import *
# import xlrd

#
# try:
#     __author__ = 'Unemi'
#
#     title = easyxf(
#         'font: name Times New Roman, color-index black, bold on , height 220; alignment: horiz left')
#     title2 = easyxf(
#         'font: name Verdana, color-index black, bold on , height 170; alignment: horiz left')
#
#     fuentecabecera = easyxf(
#         'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
#     fuentenormal = easyxf(
#         'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
#     fuentenormalneg = easyxf(
#         'font: name Verdana, color-index black, bold on , height 150; borders: left thin, right thin, top thin, bottom thin')
#     fuentenormalder = easyxf(
#         'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right')
#     fuentemoneda = easyxf(
#         'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right',
#         num_format_str=' "$" #,##0.00')
#     fuentemonedaneg = easyxf(
#         'font: name Verdana, color-index black, bold on, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right; pattern: pattern solid, fore_colour gray25',
#         num_format_str=' "$" #,##0.00')
#
#     fuentenumero = easyxf(
#         'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right',
#         num_format_str='#,##0.00')
#
#     font_style = XFStyle()
#     font_style.font.bold = True
#     font_style2 = XFStyle()
#     font_style2.font.bold = False
#     wb = Workbook(encoding='utf-8')
#     # ws = wb.add_sheet('Listado')
#
#     output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media', 'maestriainformes'))
#     nombre = "COHORTES_" + datetime.now().strftime('%Y%m%d_%H%M%S') + ".xls"
#     filename = os.path.join(output_folder, nombre)
#     hoy = datetime.now().date()
#     listaperiodos = Materia.objects.values_list('nivel__periodo_id').filter(inicio__gte=hoy, nivel__periodo__tipo_id__in=[3, 4]).distinct()
#     periodosposgrado = Periodo.objects.filter(tipo_id__in=[3, 4], pk__in=listaperiodos).order_by('nombre')
#     print(periodosposgrado.count())
#     for lisper in periodosposgrado:
#         ws = wb.add_sheet(lisper.nombre[0:23]+'_'+str(lisper.id))
#
#         ws.write_merge(0, 0, 0, 6, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
#         ws.write_merge(2, 2, 0, 6, lisper.nombre, title)
#         row_num = 5
#
#         columns = [
#             (u"MALLA", 12000),
#             (u"ASIGNATURA", 8300),
#             (u"TIENE DISTRIBUTIVO", 8300),
#             (u"CODIGOASIGNATURAMALLA", 2500)
#         ]
#         for col_num in range(len(columns)):
#             ws.write(row_num, col_num, columns[col_num][0], fuentecabecera)
#             ws.col(col_num).width = columns[col_num][1]
#         date_format = xlwt.XFStyle()
#         date_format.num_format_str = 'yyyy/mm/dd'
#
#         carreras = Carrera.objects.values_list('id').filter(pk__in=Materia.objects.values_list('asignaturamalla__malla__carrera').filter(inicio__gte=hoy,  nivel__periodo_id=lisper.id,   asignaturamalla__malla__carrera__coordinacion=7).distinct()).distinct()
#         mallas = Malla.objects.values_list('id').filter(pk__in=Materia.objects.values_list('asignaturamalla__malla').filter(nivel__periodo_id=lisper.id, asignaturamalla__malla__carrera_id__in=carreras).distinct()).distinct()
#         listamaterias = AsignaturaMalla.objects.filter(malla_id__in=mallas, status=True).order_by('malla__carrera', 'malla__inicio')
#         for lista in listamaterias:
#
#             row_num += 1
#             tiene = 'NO'
#             if lista.materia_set.values("id").filter(nivel__periodo_id=lisper.id, status=True).exists():
#                 tiene = 'SI'
#             ws.write(row_num, 0, lista.malla.carrera.nombre + '(año' + str(lista.malla.inicio) + ')', fuentenormalder)
#             ws.write(row_num, 1, lista.asignatura.nombre, fuentenormalder)
#             ws.write(row_num, 2, tiene, fuentenormalder)
#             ws.write(row_num, 3, lista.id, fuentenormalder)
#         wb.save(filename)
# except Exception as ex:
#     pass




# #
# from django.http import HttpResponse
# from xlwt import *
#
# try:
#     __author__ = 'Unemi'
#     title = easyxf('font: name Times New Roman, color-index black, bold on , height 220; alignment: horiz left')
#     title2 = easyxf('font: name Verdana, color-index black, bold on , height 170; alignment: horiz left')
#     fuentecabecera = easyxf('font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
#     fuentenormal = easyxf('font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
#     fuentenormalneg = easyxf('font: name Verdana, color-index black, bold on , height 150; borders: left thin, right thin, top thin, bottom thin')
#     fuentenormalder = easyxf('font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right')
#     fuentemoneda = easyxf('font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right', num_format_str=' "$" #,##0.00')
#     fuentemonedaneg = easyxf('font: name Verdana, color-index black, bold on, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right; pattern: pattern solid, fore_colour gray25', num_format_str=' "$" #,##0.00')
#     fuentenumero = easyxf('font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right', num_format_str='#,##0.00')
#
#     font_style = XFStyle()
#     font_style.font.bold = True
#     font_style2 = XFStyle()
#     font_style2.font.bold = False
#     wb = Workbook(encoding='utf-8')
#     # ws = wb.add_sheet('Listado')
# #
#     output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media', 'maestriainformes'))
#     nombre = "LISTADOPREFERENCIAS_" + datetime.now().strftime('%Y%m%d_%H%M%S') + ".xls"
#     filename = os.path.join(output_folder, nombre)
#     hoy = datetime.now().date()
#     listaperiodos = Materia.objects.values_list('nivel__periodo_id').filter(inicio__gte=hoy, nivel__periodo__tipo_id__in=[3, 4]).distinct()
#     # periodosposgrado = Periodo.objects.filter(tipo_id__in=[3, 4], pk__in=listaperiodos).order_by('nombre')
#     # print(periodosposgrado.count())
#     ws = wb.add_sheet('LISTADO')
#     # ws.write_merge(0, 0, 0, 6, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
#     # ws.write_merge(2, 2, 0, 6, 'listado', title)
#     row_num = 2
#
#     columns = [
#         (u"DOCENTE", 12000),
#         (u"TITULAR", 2500),
#         (u"ADMISION 1", 2500),
#         (u"ADMISION 2", 2500),
#         (u"SEMESTRE 1", 2500),
#         (u"SEMESTRE 2", 2500),
#         (u"POSGRADO", 2500),
#         (u"", 2500)
#     ]
#     for col_num in range(len(columns)):
#         ws.write(row_num, col_num, columns[col_num][0], fuentecabecera)
#         ws.col(col_num).width = columns[col_num][1]
#     date_format = xlwt.XFStyle()
#     date_format.num_format_str = 'yyyy/mm/dd'
#
#     listadistributivo = ProfesorDistributivoHoras.objects.filter(periodo_id=90, status=True).exclude(profesor__persona__apellido1__icontains='APELLIDOVIRTUAL').exclude(profesor__persona__nombres__icontains='POR DEFINIR').exclude(profesor__persona__apellido1__icontains='POR DEFINIR').exclude(profesor__persona__apellido2__icontains='POR DEFINIR').order_by('profesor__persona__apellido1', 'profesor__persona__apellido2')
#     for lista in listadistributivo:
#         row_num += 1
#         docente = lista.profesor.persona.apellido1 + ' ' + lista.profesor.persona.apellido2 + ' ' + lista.profesor.persona.nombres
#         admision1 = 'NO'
#         if lista.profesor.asignaturamallapreferencia_set.filter(periodo_id=95, status=True).exists():
#             admision1 = 'SI'
#         admision2 = 'NO'
#         if lista.profesor.asignaturamallapreferencia_set.filter(periodo_id=99, status=True).exists():
#             admision2 = 'SI'
#         semestre1 = 'NO'
#         if lista.profesor.asignaturamallapreferencia_set.filter(periodo_id=96, status=True).exists():
#             semestre1 = 'SI'
#         semestre2 = 'NO'
#         if lista.profesor.asignaturamallapreferencia_set.filter(periodo_id=97, status=True).exists():
#             semestre2 = 'SI'
#         posgrado = 'NO'
#         if lista.profesor.asignaturamallapreferenciaposgrado_set.filter(status=True).exists():
#             posgrado = 'SI'
#         aplica = 'SI APLICA'
#         if admision1=='NO':
#             if admision2 =='NO':
#                 if semestre1 =='NO':
#                     if semestre2 =='NO':
#                         if posgrado =='NO':
#                             aplica = 'NO APLICA'
#         titular = 'NO'
#         if IngresoPersonal.objects.filter(persona=lista.profesor.persona, regimenlaboral_id=2,estado=1, nombramiento=True, status=True):
#             titular='SI'
#         ws.write(row_num, 0, docente)
#         ws.write(row_num, 1, titular, fuentenormalder)
#         ws.write(row_num, 2, admision1, fuentenormalder)
#         ws.write(row_num, 3, admision2, fuentenormalder)
#         ws.write(row_num, 4, semestre1, fuentenormalder)
#         ws.write(row_num, 5, semestre2, fuentenormalder)
#         ws.write(row_num, 6, posgrado, fuentenormalder)
#         ws.write(row_num, 7, aplica, fuentenormalder)
#     wb.save(filename)
# except Exception as ex:
#     pass
#
#

# listado = AsignaturaMallaPreferencia.objects.filter(periodo__id__in=[95,96,97,99], status=True).order_by('asignaturamalla__malla__carrera__modalidad')
#
# lis = Materia.objects.values_list('asignaturamalla__malla__carrera__coordinacion__nombre', 'asignaturamalla__malla__carrera__modalidad', 'asignaturamalla__malla__carrera__coordinacion').filter(nivel__periodo__id__in=[95,96,97,99], status=True).exclude(asignaturamalla__malla__carrera__coordinacion=9).distinct()
# lis1 = Materia.objects.values_list('asignaturamalla__malla__carrera__coordinacionvalida__nombre', 'asignaturamalla__malla__carrera__modalidad', 'asignaturamalla__malla__carrera__coordinacionvalida__id').filter(nivel__periodo__id__in=[95,96,97,99],asignaturamalla__malla__carrera__coordinacion=9, status=True).distinct()
# cadena1 = []
# for li in lis:
#     cadena1.append([li[0],li[1],li[2]])
# for li1 in lis1:
#     if li1[0]:
#         noingresar = 0
#         for cade in cadena1:
#             if cade[0]==li1[0]:
#                 if cade[1]==li1[1]:
#                     if cade[2]==li1[2]:
#                         noingresar = 1
#         if noingresar == 0:
#             cadena1.append([li1[0],li1[1],li1[2]])
# cadena1.sort(key=lambda coordinacion: coordinacion[0])
# listado = listado2 = AsignaturaMallaPreferencia.objects.filter(periodo__id__in=[95,96,97,99], status=True).order_by('asignaturamalla__malla__carrera__modalidad')
# ResultadoPreferencias.objects.filter(status=True).delete()
# for lis in cadena1:
#     listaadm1 = ProfesorMateria.objects.values_list('profesor_id', 'materia__asignaturamalla_id').filter(
#         materia__asignaturamalla__malla__carrera__modalidad=lis[1],
#         materia__asignaturamalla__malla__carrera__coordinacion=lis[2],
#         materia__nivel__periodo_id=95).distinct()
#     cuentaadm1 = 0
#     for comp in listaadm1:
#         semesuno = listado2.filter(profesor_id=comp[0], asignaturamalla_id=comp[1],
#                                    asignaturamalla__malla__carrera__modalidad=lis[1],
#                                    asignaturamalla__malla__carrera__coordinacion=lis[2], periodo__id=95,
#                                    status=True).distinct()
#         if semesuno:
#             cuentaadm1 += 1
#     if cuentaadm1 == 0:
#         calculo = 0
#     else:
#         calculo = round((cuentaadm1 * 100) / listaadm1.count(), 2)
#     listaadm2 = ProfesorMateria.objects.values_list('profesor_id', 'materia__asignaturamalla_id').filter(
#         materia__asignaturamalla__malla__carrera__modalidad=lis[1],
#         materia__asignaturamalla__malla__carrera__coordinacion=lis[2],
#         materia__nivel__periodo_id=99).distinct()
#     cuentaadm2 = 0
#     for comp in listaadm2:
#         semesuno = listado2.filter(profesor_id=comp[0], asignaturamalla_id=comp[1],
#                                    asignaturamalla__malla__carrera__modalidad=lis[1],
#                                    asignaturamalla__malla__carrera__coordinacion=lis[2], periodo__id=99,
#                                    status=True).distinct()
#         if semesuno:
#             cuentaadm2 += 1
#     if cuentaadm2 == 0:
#         calculo2 = 0
#     else:
#         calculo2 = round((cuentaadm2 * 100) / listaadm2.count(), 2)
#
#
#     listaadm3 = ProfesorMateria.objects.values_list('profesor_id', 'materia__asignaturamalla_id').filter(materia__asignaturamalla__malla__carrera__modalidad=lis[1],
#                                                                                          materia__asignaturamalla__malla__carrera__coordinacion=lis[2],
#                                                                                          materia__nivel__periodo_id=96).distinct()
#     cuentaadm3 = 0
#     for comp in listaadm3:
#         semesuno = listado2.filter(profesor_id=comp[0], asignaturamalla_id=comp[1], asignaturamalla__malla__carrera__modalidad=lis[1],asignaturamalla__malla__carrera__coordinacion=lis[2],periodo__id=96, status=True).distinct()
#         if semesuno:
#             cuentaadm3 += 1
#     if cuentaadm3 == 0:
#         calculo3 = 0
#     else:
#         calculo3 = round((cuentaadm3 * 100) / listaadm3.count(), 2)
#         # SEMESTRE 2
#
#     listaadm4 = ProfesorMateria.objects.values_list('profesor_id', 'materia__asignaturamalla_id').filter(
#         materia__asignaturamalla__malla__carrera__modalidad=lis[1],
#         materia__asignaturamalla__malla__carrera__coordinacion=lis[2],
#         materia__nivel__periodo_id=97).distinct()
#     cuentaadm4 = 0
#     for comp in listaadm3:
#         semesuno = listado2.filter(profesor_id=comp[0], asignaturamalla_id=comp[1],
#                                    asignaturamalla__malla__carrera__modalidad=lis[1],
#                                    asignaturamalla__malla__carrera__coordinacion=lis[2], periodo__id=97,
#                                    status=True).distinct()
#         if semesuno:
#             cuentaadm4 += 1
#     if cuentaadm4 == 0:
#         calculo4 = 0
#     else:
#         calculo4 = round((cuentaadm4 * 100) / listaadm4.count(), 2)
#     print(str(lis[2]) + ',' + str(lis[0]) + ',' + str(lis[1]) + ',(' + str(listaadm1.count()) + ',' + str(cuentaadm1) + ',' + str(calculo) +')' + ',(' + str(listaadm2.count()) + ',' + str(cuentaadm2) + ',' + str(calculo2) +')'+ ',(' + str(listaadm3.count()) + ',' + str(cuentaadm3) + ',' + str(calculo3) +')'+ ',(' + str(listaadm4.count()) + ',' + str(cuentaadm4) + ',' + str(calculo4) +')')
#
#     resultado = ResultadoPreferencias(coordinacion_id=lis[2],
#                                       modalidad_id=lis[1],
#                                       admisionprimer=calculo,
#                                       admisionsegundo=calculo2,
#                                       semestreprimer=calculo3,
#                                       semestresegundo=calculo4)
#     resultado.save()








from django.http import HttpResponse
# from xlwt import *
#
# try:
#     __author__ = 'Unemi'
#     title = easyxf('font: name Times New Roman, color-index black, bold on , height 220; alignment: horiz left')
#     title2 = easyxf('font: name Verdana, color-index black, bold on , height 170; alignment: horiz left')
#     fuentecabecera = easyxf('font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
#     fuentenormal = easyxf('font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
#     fuentenormalneg = easyxf('font: name Verdana, color-index black, bold on , height 150; borders: left thin, right thin, top thin, bottom thin')
#     fuentenormalder = easyxf('font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right')
#     fuentemoneda = easyxf('font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right', num_format_str=' "$" #,##0.00')
#     fuentemonedaneg = easyxf('font: name Verdana, color-index black, bold on, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right; pattern: pattern solid, fore_colour gray25', num_format_str=' "$" #,##0.00')
#     fuentenumero = easyxf('font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin; alignment: horiz right', num_format_str='#,##0.00')
#
#     font_style = XFStyle()
#     font_style.font.bold = True
#     font_style2 = XFStyle()
#     font_style2.font.bold = False
#     wb = Workbook(encoding='utf-8')
#     # ws = wb.add_sheet('Listado')
# #
#     output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media', 'maestriainformes'))
#     nombre = "listadoporcntajesmaterias" + datetime.now().strftime('%Y%m%d_%H%M%S') + ".xls"
#     filename = os.path.join(output_folder, nombre)
#     hoy = datetime.now().date()
#     listadoperiodo = Periodo.objects.filter(pk__in=[95,96,97,99], status=True)
#     for lisper in listadoperiodo:
#         ws = wb.add_sheet(lisper.nombre[0:23] + '_' + str(lisper.id))
#         ws.write_merge(0, 0, 0, 6, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
#         listadodistributivomaterias = Materia.objects.values_list('asignaturamalla__malla__carrera_id', 'asignaturamalla__malla__carrera__nombre').filter(nivel__periodo_id=lisper.id, status=True).distinct()
#         row_num = 2
#
#         columns = [
#             (u"CODIGO", 2000),
#             (u"CARRERA", 15000),
#             (u"TOTAL PLANIFICADAS", 3000),
#             (u"TOTAL CON DOCENTE", 3000),
#             (u"PORCENTAJE", 3000),
#             (u"TOTAL CON HORARIO", 3000),
#             (u"PORCENTAJE HORARIO", 3000)
#         ]
#         for col_num in range(len(columns)):
#             ws.write(row_num, col_num, columns[col_num][0], fuentecabecera)
#             ws.col(col_num).width = columns[col_num][1]
#         date_format = xlwt.XFStyle()
#         date_format.num_format_str = 'yyyy/mm/dd'
#         sumaporcentaje = 0
#         sumaporcentajehorario = 0
#         porcentajetotal = 0
#         porcentajetotalhorario = 0
#         for lista in listadodistributivomaterias:
#             listaasignaturas = Materia.objects.values_list('id').filter(asignaturamalla__malla__carrera_id=lista[0], nivel__periodo_id=lisper.id, status=True)
#
#             # listaconprofesor = ProfesorMateria.objects.values_list('materia_id').filter(materia_id__in=listaasignaturas, status=True).exclude(profesor__persona__apellido1__icontains='DEFINIR').exclude(profesor__persona__apellido1__icontains='VIRTUAL').exclude(profesor__persona__apellido1__icontains='APELLIDOVIRT').exclude(profesor__persona__nombres__icontains='DEFINIR').exclude(profesor__persona__nombres__icontains='VIRTUAL').exclude(profesor__persona__apellido2__icontains='VIRTUAL').distinct()
#             listaconprofesor = ProfesorMateria.objects.values_list('materia_id').filter(materia_id__in=listaasignaturas, status=True).distinct()
#             listaconhorario = Clase.objects.values_list('materia_id').filter(materia_id__in=listaasignaturas, status=True).distinct()
#
#             row_num += 1
#             campo1 = lista[0]
#             campo2 = lista[1]
#             campo3 = listaasignaturas.count()
#             campo4 = listaconprofesor.count()
#             campo5 = round((listaconprofesor.count() * 100) / listaasignaturas.count(),2)
#             campo6 = listaconhorario.count()
#             campo7 = round((listaconhorario.count() * 100) / listaasignaturas.count(),2)
#
#             sumaporcentaje += campo5
#             sumaporcentajehorario += campo7
#             ws.write(row_num, 0, campo1)
#             ws.write(row_num, 1, campo2, fuentenormalder)
#             ws.write(row_num, 2, campo3, fuentenormalder)
#             ws.write(row_num, 3, campo4, fuentenormalder)
#             ws.write(row_num, 4, campo5, fuentenormalder)
#             ws.write(row_num, 5, campo6, fuentenormalder)
#             ws.write(row_num, 6, campo7, fuentenormalder)
#         row_num += 2
#         porcentajetotal = round((sumaporcentaje / listadodistributivomaterias.count()), 2)
#         porcentajetotalhorario = round((sumaporcentajehorario / listadodistributivomaterias.count()), 2)
#         ws.write(row_num, 0, '')
#         ws.write(row_num, 1, '', fuentenormalder)
#         ws.write(row_num, 2, '', fuentenormalder)
#         ws.write(row_num, 3, '', fuentenormalder)
#         ws.write(row_num, 4, porcentajetotal, fuentenormalder)
#         ws.write(row_num, 5, '', fuentenormalder)
#         ws.write(row_num, 6, porcentajetotalhorario, fuentenormalder)
#
#
#
#
#         listadodistributivomateriascoordinacion = Materia.objects.values_list('asignaturamalla__malla__carrera__coordinacion__id', 'asignaturamalla__malla__carrera__coordinacion__nombre').filter( nivel__periodo_id=lisper.id, status=True).distinct()
#
#         columns = [
#             (u"CODIGO", 2000),
#             (u"COORDINACION", 15000),
#             # (u"TOTAL PLANIFICADAS", 3000),
#             # (u"TOTAL CON DOCENTE", 3000),
#             (u"PORCENTAJE", 3000),
#             # (u"TOTAL CON HORARIO", 3000),
#             (u"PORCENTAJE HORARIO", 3000)
#         ]
#         row_num += 1
#         for col_num in range(len(columns)):
#             ws.write(row_num, col_num, columns[col_num][0], fuentecabecera)
#             ws.col(col_num).width = columns[col_num][1]
#         sumaporcentaje = 0
#         porcentajetotal = 0
#         for lista in listadodistributivomateriascoordinacion:
#             row_num += 1
#             listaasignaturascoor = Materia.objects.values_list('id').filter(asignaturamalla__malla__carrera__coordinacion__id=lista[0], nivel__periodo_id=lisper.id, status=True)
#             # listaconprofesorcoor = ProfesorMateria.objects.values_list('materia_id').filter(materia_id__in=listaasignaturascoor, status=True).exclude( profesor__persona__apellido1__icontains='DEFINIR').exclude(
#             #     profesor__persona__apellido1__icontains='VIRTUAL').exclude(
#             #     profesor__persona__apellido1__icontains='APELLIDOVIRT').exclude(
#             #     profesor__persona__nombres__icontains='DEFINIR').exclude(
#             #     profesor__persona__nombres__icontains='VIRTUAL').exclude(
#             #     profesor__persona__apellido2__icontains='VIRTUAL').distinct()
#             listaconprofesorcoor = ProfesorMateria.objects.values_list('materia_id').filter(materia_id__in=listaasignaturascoor, status=True).distinct()
#             listaconhorariocoor = Clase.objects.values_list('materia_id').filter(materia_id__in=listaasignaturascoor, status=True).distinct()
#
#             campo1 = lista[0]
#             campo2 = lista[1]
#             campo3 = listaasignaturascoor.count()
#             campo4 = listaconprofesorcoor.count()
#             campo5 = round((listaconprofesorcoor.count() * 100) / listaasignaturascoor.count(), 2)
#             # campo6 = listaconhorario.count()
#             campo7 = round((listaconhorariocoor.count() * 100) / listaasignaturascoor.count(), 2)
#             #
#             # sumaporcentaje += campo5
#             #
#             ws.write(row_num, 0, campo1)
#             ws.write(row_num, 1, campo2, fuentenormalder)
#             # ws.write(row_num, 2, campo3, fuentenormalder)
#             # ws.write(row_num, 3, campo4, fuentenormalder)
#             ws.write(row_num, 2, campo5, fuentenormalder)
#             # ws.write(row_num, 5, campo6, fuentenormalder)
#             ws.write(row_num, 3, campo7, fuentenormalder)
#         # row_num += 2
#         # porcentajetotal = round((sumaporcentaje / listadodistributivomaterias.count()), 2)
#         # ws.write(row_num, 0, '')
#         # ws.write(row_num, 1, '', fuentenormalder)
#         # ws.write(row_num, 2, '', fuentenormalder)
#         # ws.write(row_num, 3, '', fuentenormalder)
#         # ws.write(row_num, 4, porcentajetotal, fuentenormalder)
#         # ws.write(row_num, 5, '', fuentenormalder)
#         # ws.write(row_num, 6, '', fuentenormalder)
#
#         wb.save(filename)
# except Exception as ex:
#     pass


#
#
# import xlsxwriter
# import io
# cursor = connection.cursor()
# row_ = 0
# col_ = 0
# idcoordinacion = 4
# idencuesta = 44
# idpregunta = 428
# cadena = [44,44]
# # output = io.BytesIO()
# # workbook = xlsxwriter.Workbook(output)
# # add worksheet to workbook
# # worksheet = workbook.add_worksheet()
#
# # print(chart_data)
# xls_file = 'ejemplo.xlsx'
# workbook = xlsxwriter.Workbook(xls_file)
# # add worksheet to workbook
# worksheet = workbook.add_worksheet()
#
#
# sql_encuesta = """
#                 SELECT distinct pregu.id,pregu.nombre
#                 FROM sga_sagpreguntaencuesta encupre,
#                 sga_sagencuesta encu,sga_sagpregunta pregu
#                 WHERE encupre.sagencuesta_id=encu.id
#                 AND encupre.sagpregunta_id=pregu.id
#                 AND encu.id=44
#
#                 ORDER BY pregu.id
#
#                 """
# cursor.execute(sql_encuesta)
# recorrepregunta = cursor.fetchall()
# for pregunta in  recorrepregunta:
#     idpregunta = pregunta[0]
#     nompregunta = pregunta[1]
#
#     # for indice in cadena:
#     chart_data = []
#     sql_mensajes = """
#                     SELECT det.valor as escala,COUNT(pregu.id) as total
#                     FROM sga_sagresultadoencuestadetalle det,sga_sagresultadoencuesta resul,
#                     sga_inscripcion ins,sga_coordinacion coor,sga_sagpreguntaencuesta pren,
#                     sga_sagencuesta encu,sga_sagperiodo sagperi,sga_sagpregunta pregu
#                     WHERE det.sagresultadoencuesta_id=resul.id
#                     AND resul.inscripcion_id=ins.id
#                     AND ins.coordinacion_id=coor.id
#                     AND det.preguntaencuesta_id=pren.id
#                     AND pren.sagencuesta_id=encu.id
#                     AND encu.sagperiodo_id=sagperi.id
#                     AND pren.sagpregunta_id=pregu.id
#                     AND coor.id=%s
#                     AND encu.id=%s
#                     AND pregu.id=%s
#                     AND encu."status"=TRUE
#                     GROUP BY encu.nombre,pregu.id,pregu.nombre,det.valor
#                     """ % (idcoordinacion,idencuesta, int(idpregunta))
#     cursor.execute(sql_mensajes)
#     mensajes_leidos = cursor.fetchall()
#     ubicacionchart = mensajes_leidos.__len__() + 3
#     totalencuestados = 0
#     print(mensajes_leidos)
#     for indicetotal in  mensajes_leidos:
#         totalencuestados = totalencuestados + indicetotal[1]
#     totalencuestados = totalencuestados
#     for mensaje in  mensajes_leidos:
#         chart_data.append({
#         'ESCALA': mensaje[0],
#         'TOTAL': mensaje[1],
#         'PORCENTAJE': round(mensaje[1] * 100 / totalencuestados,2)})
#     merge_format = workbook.add_format({
#         'bold': 1,
#         'border': 1,
#         'align': 'left',
#         'valign': 'vcenter'
#         # 'fg_color': 'yellow'
#     })
#
#     row_ += 1
#     worksheet.merge_range('A' + str(row_) + ':H' + str(row_), nompregunta, merge_format)
#     worksheet.write(row_, col_, 'ESCALA',merge_format)
#     col_ += 1
#     worksheet.write(row_, col_, 'TOTAL',merge_format)
#     col_ += 1
#     worksheet.write(row_, col_, 'PORCENTAJE',merge_format)
#     row_ += 1
#     leedatos = row_ + 1
#     print(row_)
#     for item in chart_data:
#         col_ = 0
#         worksheet.write(row_, col_, item['ESCALA'])
#         col_ += 1
#         worksheet.write(row_, col_, item['TOTAL'])
#         col_ += 1
#         worksheet.write(row_, col_, item['PORCENTAJE'])
#         row_ += 1
#
#     column_chart = workbook.add_chart({'type': 'column'})
#
#     column_chart.add_series({
#         'name': 'Porcentaje',
#         'categories': '=Sheet1!$A$'+str(leedatos)+':$A$%s' % row_,
#         'values': '=Sheet1!$C$'+str(leedatos)+':$C$%s' % row_,
#         'data_labels': {'value': True, 'leader_lines': True},
#         'marker': {'type': 'circle'}
#     })
#
#
#     # print(ubicacionchart)
#     worksheet.insert_chart('D'+str(row_ + 1), column_chart)
#     row_ += ubicacionchart + 15
#     col_ = 0
# workbook.close()
# output.seek(0)
# filename = 'gfhfgh.xlsx'
# response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
# response['Content-Disposition'] = 'attachment; filename=%s' % filename
# print(response)
# row_ = 40
# col_ = 0
#
# # write headers
# worksheet.write(row_, col_, 'NAME')
# col_ += 1
# worksheet.write(row_, col_, 'VALUE')
# col_ += 1
# worksheet.write(row_, col_, 'VALUE2')
# row_ += 1
#
# chart_data2 = [
#     {'name': 'Lorem', 'value': 23, 'value1': 23},
#     {'name': 'Ipsum', 'value': 48, 'value1': 23},
#     {'name': 'Dolor', 'value': 15, 'value1': 23},
#     {'name': 'Sit', 'value': 8, 'value1': 23},
#     {'name': 'Amet', 'value': 32, 'value1': 23}
# ]
# # write sample data
# for item in chart_data2:
#     col_ = 0
#     worksheet.write(row_, col_, item['name'])
#     col_ += 1
#     worksheet.write(row_, col_, item['value'])
#     col_ += 1
#     worksheet.write(row_, col_, item['value1'])
#     row_ += 1
# # create pie chart
#
# #
# # # create column chart
# column_chart1 = workbook.add_chart({'type': 'column'})
# #
# # add serie to column chart
# column_chart1.add_series({
#     'name': 'Porcentaje profesor',
#     'categories': '=Sheet1!$A$42:$A$%s' % row_,
#     'values': '=Sheet1!$B$42:$B$%s' % row_,
#     'data_labels': {'value': True, 'leader_lines': True},
#     'marker': {'type': 'circle'}
# })
# worksheet.insert_chart('D40', column_chart1)
# workbook.close()
#

#
# materias = Materia.objects.filter(nivel__periodo_id=89, status=True).exclude(pk__in=[21962,21969,21953,21957,23906,23907,21968,21961,21956,21952,21967,21960])
# print(materias.count())
#
# for mat in materias:
#     mat.nivel_id=494
#     mat.save()
#     print('actualizado')
# #
# matri = Matricula.objects.values_list('id').filter(pk=193642, materiaasignada__materia_id__in=[21962,21969,21953,21957,23906,23907,21968,21961,21956,21952,21967,21960], status=True).distinct()
#
# for lismatri in matri:
#     listadomateriaasignada = MateriaAsignada.objects.filter(matricula_id=lismatri[0], status=True).exclude(materia_id__in=[21962,21969,21953,21957,23906,23907,21968,21961,21956,21952,21967,21960])
#     matrialumano = Matricula.objects.get(pk=lismatri[0])
#     matrialumano.pk = None
#     matrialumano.nivel_id = 494
#     matrialumano.save()
#     for lista in listadomateriaasignada:
#         lista.matricula=matrialumano
#         lista.save()
#     print(matrialumano.id)
#
# print(matri.count())

# materias = Materia.objects.filter(pk__in=[21965,21973,21972,21971,21970,21966,21964,21963,21959,21958,21955,21954], asignaturamalla__malla__carrera_id=1, status=True)
#
# print(materias.count())
#
# matriculas = MateriaAsignada.objects.values_list('matricula_id','materia__paralelo','matricula__inscripcion_id').filter(materia__asignaturamalla__malla__carrera_id=1, materia_id__in=[21962,21969,21953,21957,23906,23907,21968,21961,21956,21952,21967,21960], status=True).distinct()
# periodo = Periodo.objects.get(pk=96)
# for matri in matriculas:
#     inscripcion = Inscripcion.objects.get(pk=matri[2])
#     if not inscripcion.matricula_periodo(periodo):
#         matricula = Matricula(inscripcion=inscripcion,
#                               nivel_id=494,
#                               pago=False,
#                               iece=False,
#                               becado=False,
#                               porcientobeca=0,
#                               fecha=datetime.now().date(),
#                               hora=datetime.now().time(),
#                               fechatope=datetime.now().date())
#         matricula.save()
#     else:
#         matricula = Matricula.objects.get(inscripcion=inscripcion, nivel_id=494)
#     for mat in materias:
#         if mat.paralelo == matri[1]:
#             if not MateriaAsignada.objects.values('id').filter(matricula=matricula, materia=mat).exists():
#                 matriculas = matricula.inscripcion.historicorecordacademico_set.values('id').filter(asignatura=mat.asignatura, fecha__lt=mat.nivel.fin).count() + 1
#                 materiaasignada = MateriaAsignada(matricula=matricula,
#                                                   materia=mat,
#                                                   notafinal=0,
#                                                   asistenciafinal=0,
#                                                   cerrado=False,
#                                                   matriculas=matriculas,
#                                                   observaciones='',
#                                                   estado_id=NOTA_ESTADO_EN_CURSO)
#                 materiaasignada.save()
#                 materiaasignada.asistencias()
#                 materiaasignada.evaluacion()
#                 materiaasignada.mis_planificaciones()
#                 materiaasignada.save()
#     matricula.actualizar_horas_creditos()
#     matricula.estado_matricula = 2
#     matricula.save()
#     matricula.calcula_nivel()
#     inscripcion.actualizar_nivel()
#     codigomatricula = Matricula.objects.get(pk=matri[0])
#     if not matricula.matriculagruposocioeconomico_set.filter(status=True):
#         gruposocio = MatriculaGrupoSocioEconomico.objects.get(matricula=codigomatricula, status=True)
#         gruposocio.pk = None
#         gruposocio.matricula = matricula
#         gruposocio.save()
#         print(matricula)
# listadorubricas = Rubrica.objects.filter(tipoprofesor_id=12)
# a =  0
# for ru in listadorubricas:
#     a = a + 1
#     ru.delete()
#     print(a)
# listadorubricas = Rubrica.objects.filter(pk=464)
# for ru in  listadorubricas:
#     rubrica = Rubrica.objects.get(pk=ru.id)
#     listacriterios = rubrica.rubricacriteriodocencia_set.filter(status=True)
#     listadocaracteristicas = rubrica.rubricacaracteristica_set.filter(status=True)
#     listadopreguntas = rubrica.rubricapreguntas_set.order_by('orden')
#     nuevarubrica = ru
#     nuevarubrica.pk = None
#     nuevarubrica.proceso_id = 91
#     nuevarubrica.save()
#     print(ru.id)
#     for cri in listacriterios:
#         print(cri.id)
#         nuevacri = cri
#         nuevacri.pk = None
#         nuevacri.rubrica = nuevarubrica
#         nuevacri.save()
#     for carac in listadocaracteristicas:
#         nuevacarac = carac
#         nuevacarac.pk = None
#         nuevacarac.rubrica = nuevarubrica
#         nuevacarac.save()
#     for preg in listadopreguntas:
#         nuevapreg = preg
#         nuevapreg.pk = None
#         nuevapreg.rubrica = nuevarubrica
#         nuevapreg.save()
# print(listadorubricas.count())
#
#
# profesores = ProfesorDistributivoHoras.objects.filter(periodo_id=112, status=True).exclude(coordinacion_id=9)
#
# for listado in profesores:
#     profesorhora = ProfesorConfigurarHoras(periodo_id=119,
#                                            profesor=listado.profesor,
#                                            dedicacion=listado.dedicacion)
#     profesorhora.save()
# profesor = Profesor.objects.get(pk=1106)
# sumahoras = null_to_numeric(profesor.profesormateria_set.filter(materia__nivel__periodo_id=113,status=True).aggregate(suma=Sum('hora'))['suma'])
# print(sumahoras)
# mate = Materia.objects.get(pk=36142)
# fini=str(mate.inicio)
# finim = convertir_fecha_hora(fini + " 00:00")
# print(finim)




