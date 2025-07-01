# -*- coding: latin-1 -*-
import json
import random
from datetime import datetime, timedelta

from openpyxl import workbook as openxl
from openpyxl.styles import Font as openxlFont
from openpyxl.styles.alignment import Alignment as alin

from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Q
from django.forms import model_to_dict
from django.http import JsonResponse, HttpResponseRedirect, HttpResponse
from django.shortcuts import render
from django.template.loader import get_template
from django.utils.text import capfirst

from decorators import secure_module, last_access
from empleo.models import OfertaLaboralEmpresa, ESTADOS_APLICACION, TerminosAplicaOferta, OfertaTermino, PersonaAplicaOferta, NIVEL_INSTRUCCION, MODALIDAD, DEDICACION, JORNADA, TIEMPO,OPCION_OBSERVACION_HOJAVIDA, OPCION_OBSERVACION_CONTRATO
from empleo.postular import validar_campos
from empresa.forms import RepresentanteForm, OfertaLaboralForm, GestionarContratoForm, ObservacionHojaVidaForm
from empresa.models import RepresentantesEmpresa
from settings import EMAIL_DOMAIN
from sga.commonviews import adduserdata
from sga.funciones import MiPaginador, puede_realizar_accion, log, notificacion, variable_valor
from sga.models import Empleador, Persona, Carrera, Inscripcion, Graduado, CarreraGruposCarrera, Titulacion, Titulo, \
    CertificadoIdioma, NivelTitulacion, CamposTitulosPostulacion, Capacitacion, CUENTAS_CORREOS, Notificacion
from sga.tasks import send_html_mail
from sga.templatetags.sga_extras import encrypt


@login_required(redirect_field_name='ret', login_url='empresa/loginempresa')
# @secure_module
@last_access
@transaction.atomic()
def view(request):
    data = {}
    adduserdata(request, data)
    data['url_'] = request.path
    perfilprincipal = request.session['perfilprincipal']
    persona = request.session['persona']
    data['hoy'] = hoy = datetime.now().date()
    data['currenttime'] = datetime.now()
    data['perfil'] = persona.mi_perfil()
    data['empresa'] = empresa = Empleador.objects.get(persona=persona)


    if request.method == 'POST':
        action = request.POST['action']
        if action == 'add':
            try:
                    data['hoy'] = datetime.now().date()
                # f = OfertaLaboralForm(request.POST)
                # f.set_required_false(request.POST, request.POST.getlist('carrera'), request.POST.getlist('titulorquerido'))
                # if f.is_valid():
                    finicio= request.POST['finicio']
                    ffin= request.POST['ffin']
                    finiciopostulacion= request.POST['finiciopostulacion']
                    ffinpostulacion= request.POST['ffinpostlacion']
                    finiciorevision = request.POST['finiciorevision']
                    ffinrevision = request.POST['ffinrevision']

                    valida = OfertaLaboralEmpresa.objects.filter(encargado=request.POST['encargado'], titulo=request.POST['titulo'],
                                                           finicio=request.POST['finicio'], ffin=request.POST['ffin'])
                    if 'nivel' in request.POST:
                        valida = valida.filter(nivel=request.POST['nivel'])
                    if valida.exists():
                        raise NameError('Ya existe una oferta con estos datos')
                    if ffin <= finicio:
                        raise NameError('La fecha final del proceso debe ser mayor a la fecha de inicio del proceso')
                    # FECHAS DE POSTULACION
                    if finiciopostulacion < finicio or finiciopostulacion > ffin:
                        raise NameError('La fecha de inicio de postulacion debe estar dentro del rango de las fechas de proceso')
                    if ffinpostulacion < finicio or ffinpostulacion > ffin:
                        raise NameError('La fecha de finalizacion de postulacion debe estar dentro del rango de las fechas de proceso')
                    if ffinpostulacion <= finiciopostulacion:
                        raise NameError('La fecha fin de postulacion debe ser mayor a la fecha de inicio de la postulacion')

                    if finiciorevision < ffinpostulacion or finiciorevision > ffin:
                        raise NameError('La fecha de inicio de revision debe estar dentro del rango de las fechas de proceso')
                    if ffinrevision < finiciopostulacion or ffinrevision > ffin:
                        raise NameError('La fecha de fin de revision debe estar dentro del rango de las fechas de proceso')
                    if ffinrevision <= finiciorevision:
                        raise NameError('La fecha fin de revision debe ser mayor a la fecha de inicio de la revision')
                    # carreraselect = Carrera.objects.get(id=request.POST['carrera'].id)
                    # carrerasrelacionadas = Carrera.objects.filter(status=True, areaconocimiento=carreraselect.areaconocimiento,
                    #                                               subareaconocimiento=carreraselect.subareaconocimiento,
                    #                                               subareaespecificaconocimiento=carreraselect.subareaespecificaconocimiento)
                    oferta = OfertaLaboralEmpresa(encargado_id=request.POST['encargado'],
                                           empresa=empresa,
                                           titulo=request.POST['titulo'],
                                           descripcion=request.POST['descripcion'],
                                           nivel=request.POST['nivel'] if 'nivel' in request.POST else 2,
                                           modalidad=request.POST['modalidad'],
                                           dedicacion=request.POST['dedicacion'],
                                           jornada=request.POST['jornada'],
                                           rmu=request.POST['rmu'],
                                           finicio=request.POST['finicio'],
                                           ffin=request.POST['ffin'],

                                           finiciopostulacion=request.POST['finiciopostulacion'],
                                           ffinpostlacion=request.POST['ffinpostlacion'],

                                           finiciorevision=request.POST['finiciorevision'],
                                           ffinrevision=request.POST['ffinrevision'],

                                           tipocontrato_id=request.POST['tipocontrato'],
                                           vacantes=request.POST['vacantes'],
                                           muestranombre= True if 'muestranombre' in request.POST else False,
                                           muestrarmu=True if 'muestrarmu' in request.POST else False,
                                           discapacitados=True if 'discapacitados' in request.POST else False,
                                           viajar= True if 'viajar' in request.POST else False,
                                           carropropio= True if 'carropropio' in request.POST else False,
                                           requiereexpe= True if 'requiereexpe' in request.POST else False,
                                           muestrapromedio= True if 'muestrapromedio' in request.POST else False,
                                           sexo=request.POST['sexo'],
                                           quienpostula=request.POST['quienpostula'] if 'quienpostula' in request.POST else 2,
                                           tiempoexperiencia=request.POST['tiempoexperiencia'] if 'requiereexpe' in request.POST else 4,
                                           conocimiento=request.POST['conocimiento'],
                                           funciones=request.POST['funciones'],
                                           habilidades=request.POST['habilidades'],
                                           areatrabajo=request.POST['areatrabajo'],
                                           pais_id=request.POST['pais'],
                                           provincia_id=request.POST['provincia'] if 'provincia' in request.POST else None,
                                           canton_id=request.POST['canton'] if 'canton' in request.POST else None,
                                           direccion=request.POST['direccion'])
                    if oferta.quienpostula == 1:
                        oferta.nivel = 2
                    oferta.save(request)
                    for carrera in request.POST.getlist('carrera'):
                        oferta.carrera.add(carrera)
                    for relacionadas in CarreraGruposCarrera.objects.filter(status=True, carrera__in=oferta.carrera.all().values_list('id', flat=True)).distinct():
                        for car in relacionadas.carrera.all():
                            oferta.carrerarelacionada.add(car)
                    terminos = TerminosAplicaOferta.objects.filter(status=True, activo=True)
                    for termino in terminos:
                        if termino.pk == 3 and oferta.muestrapromedio:
                            ot = OfertaTermino(oferta=oferta, termino=termino)
                        else:
                            ot = OfertaTermino(oferta=oferta, termino=termino)
                        ot.save(request)
                    send_html_mail(u"Nueva oferta laboral", "emails/nuevaofertaregistrada.html",
                                   {'sistema': u'Unemi Empresas', 'fecha': datetime.now(),
                                    'oferta': oferta, 'tit': 'Unemi - Empresa'},
                                   ['graduados@unemi.edu.ec'], [], cuenta=CUENTAS_CORREOS[17][1])
                    log(u'Adiciono oferta laboral: %s' % oferta, request, "add")
                    return JsonResponse({"result": "ok", "titulo": "Registro Guardado", "mensaje": "La oferta fue guardada correctamente. Esta oferta será revisada en las proximas horas"})
                # else:
                #     transaction.set_rollback(True)
                #     return JsonResponse({'result': "bad", "form": [{k: v[0]} for k, v in f.errors.items()],
                #                          "message": "Error en el formulario"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": str(ex)})
        elif action == 'edit':
            try:
                f = OfertaLaboralForm(request.POST)
                # if f.is_valid():
                oferta = OfertaLaboralEmpresa.objects.get(pk=request.POST['id'])
                oferta.encargado_id = request.POST['encargado']
                oferta.empresa = empresa
                oferta.titulo = request.POST['titulo']
                oferta.descripcion = request.POST['descripcion']
                oferta.nivel = request.POST['nivel'] if 'nivel' in request.POST else 2
                oferta.modalidad = request.POST['modalidad']
                oferta.dedicacion = request.POST['dedicacion']
                oferta.jornada = request.POST['jornada']
                oferta.rmu = request.POST['rmu']
                oferta.finicio = request.POST['finicio']
                oferta.ffin = request.POST['ffin']

                oferta.finiciopostulacion = request.POST['finiciopostulacion']
                oferta.ffinpostlacion = request.POST['ffinpostlacion']
                oferta.finiciorevision = request.POST['finiciorevision']
                oferta.ffinrevision = request.POST['ffinrevision']

                oferta.tipocontrato_id = request.POST['tipocontrato']
                oferta.vacantes = request.POST['vacantes']
                oferta.muestranombre = True if 'muestranombre' in request.POST else False
                oferta.muestrarmu = True if 'muestrarmu' in request.POST else False
                oferta.discapacitados = True if 'discapacitados' in request.POST else False
                oferta.viajar = True if 'viajar' in request.POST else False
                oferta.carropropio = True if 'carropropio' in request.POST else False
                oferta.requiereexpe = True if 'requiereexpe' in request.POST else False
                oferta.sexo = request.POST['sexo']
                oferta.quienpostula = request.POST['quienpostula']
                oferta.muestrapromedio = True if 'muestrapromedio' in request.POST else False
                oferta.tiempoexperiencia = request.POST['tiempoexperiencia'] if request.POST['requiereexpe'] else 4
                oferta.conocimiento = request.POST['conocimiento']
                oferta.funciones = request.POST['funciones']
                oferta.habilidades = request.POST['habilidades']
                oferta.areatrabajo = request.POST['areatrabajo']
                oferta.pais_id = request.POST['pais']
                oferta.provincia_id = request.POST['provincia'] if 'provincia' in request.POST else None
                oferta.canton_id = request.POST['canton'] if 'canton' in request.POST else None
                oferta.direccion = request.POST['direccion']
                oferta.save(request)
                oferta.carrera.clear()
                for carrera in request.POST.getlist('carrera'):
                    oferta.carrera.add(carrera)
                for relacionadas in CarreraGruposCarrera.objects.filter(status=True,
                                                                        carrera__in=oferta.carrera.all().values_list(
                                                                                'id', flat=True)).distinct():
                    for car in relacionadas.carrera.all():
                        oferta.carrerarelacionada.add(car)
                if oferta.estadooferta == 3:
                    send_html_mail(u"Oferta Laboral Corregida", "emails/ofertalaboralcorregida.html",
                                   {'sistema': u'Unemi Empresas', 'fecha': datetime.now(),
                                    'oferta': oferta, 'tit': 'Unemi - Empresa'},
                                   ['graduados@unemi.edu.ec'], [], cuenta=CUENTAS_CORREOS[17][1])
                log(u'Editó oferta laboral: %s' % oferta, request, "edit")
                return JsonResponse({"result": "ok", "titulo": "Registro Guardado", "mensaje": "La oferta fue editada correctamente"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": "Error al guardar los datos. {}".format(str(ex))})
        elif action == 'delete':
            try:
                oferta = OfertaLaboralEmpresa.objects.get(pk=int(encrypt(request.POST['id'])))
                if oferta.personaaplicaoferta_set.filter(status=True).exists():
                    return JsonResponse({"result": "bad", "mensaje": u"No puede eliminar esta oferta ya que tiene postulantes."})
                oferta.status = False
                oferta.save(request)
                log(u'Eliminó oferta laboral: %s - %s' % (oferta, oferta.empresa.nombrecorto), request, "del")
                return JsonResponse({"error": False})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": "Error al eliminar los datos. {}".format(str(ex))})
        elif action == 'detalletitulo':
            try:
                data['titulacion'] = titulacion = Titulacion.objects.get(pk=int(request.POST['id']))
                data['campostitulo'] = CamposTitulosPostulacion.objects.filter(status=True,
                                                                               titulo=titulacion.titulo)
                dettitu = titulacion.detalletitulacionbachiller_set.filter(status=True)
                data['detalletitulacionbachiller'] = dettitu.last()
                if titulacion.usuario_creacion:
                    data['personacreacion'] = Persona.objects.get(
                        usuario=titulacion.usuario_creacion) if titulacion.usuario_creacion.id > 1 else ""
                template = get_template("empleo/requisitos/modal/detalletitulo.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'html': json_content})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": "Error al obtener los datos. {}".format(str(ex))})
        elif action == 'detalleotroscertificacion':
            try:
                data['certificacion'] = certificacion = CertificadoIdioma.objects.get(pk=int(request.POST['id']))
                if certificacion.usuario_creacion:
                    data['personacreacion'] = Persona.objects.get(
                        usuario=certificacion.usuario_creacion) if certificacion.usuario_creacion.id > 1 else ""
                template = get_template("empleo/requisitos/modal/detalleidiomas.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'html': json_content})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": "Error al obtener los datos. {}".format(str(ex))})
        elif action == 'detallecapacitacion':
            try:
                data['capacitacion'] = capacitacion = Capacitacion.objects.get(pk=int(request.POST['id']))
                if capacitacion.usuario_creacion:
                    data['personacreacion'] = Persona.objects.get(usuario=capacitacion.usuario_creacion) if capacitacion.usuario_creacion.id > 1 else ""
                template = get_template("th_hojavida/detallecapacitacion.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'html': json_content})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'gestionarpostulacion':
            try:
                data['id'] = id = int(encrypt(request.POST['id']))
                postulacion = PersonaAplicaOferta.objects.get(id=id)
                postulacion.estado = int(request.POST['estado'])
                postulacion.save(request)
                estadooferta = 'No apta'

                if int(request.POST['estado']) == 2:
                    estadooferta = 'Apta'
                elif int(request.POST['estado']) == 3:
                    estadooferta = 'No apta'
                    f = ObservacionHojaVidaForm(request.POST)
                    if f.is_valid():
                        if int(f.cleaned_data['opc_hojavida']) == 0:
                            raise NameError('Debe seleccionar un motivo de rechazo')
                        postulacion.opc_hojavida = f.cleaned_data['opc_hojavida']
                        if (int(f.cleaned_data['opc_hojavida']) == 5):
                            postulacion.observacionhojavida = f.cleaned_data['observacionhojavida']
                        else:
                            postulacion.observacionhojavida = OPCION_OBSERVACION_HOJAVIDA[int(f.cleaned_data['opc_hojavida'])][1]

                        postulacion.save(request)

                notificacion = Notificacion(titulo='Postulación {}'.format(estadooferta),
                                            cuerpo='Estimad@ {} su postulación a la oferta {} fue {}'.format(postulacion.persona.nombre_completo_minus(), postulacion.oferta.titulo, estadooferta),
                                            destinatario_id=postulacion.persona.id,
                                            # departamento=departamento,
                                            url='/emp_postulaciones',
                                            object_id=postulacion.pk,
                                            prioridad=1,
                                            app_label='empleo',
                                            fecha_hora_visible=datetime.now() + timedelta(days=1)
                                            )
                notificacion.save(request)
                send_html_mail('Postulación {}'.format(estadooferta), "emails/estadopostulacionempleo.html",
                               {'sistema': u'Unemi Empleo', 'fecha': datetime.now(),
                                'postulacion': postulacion, 'tit': 'Unemi - Empleo', 'estadooferta': estadooferta},
                               postulacion.persona.lista_emails_envio(),
                               [], cuenta=CUENTAS_CORREOS[17][1])
                log(u'Gestiono postulacion: %s' % postulacion, request, "add")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                print(ex)
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": "Error en la transacción. {}".format(str(ex))})

        elif action == 'contratarapto':
            try:
                contrato = PersonaAplicaOferta.objects.get(pk=int(encrypt(request.POST['id'])))
                contrato.estcontrato = 1
                contrato.fecha_contrato = hoy
                contrato.save()
                estcontrato = 'Contratado/a'
                notificacion = Notificacion(titulo='Estado de contratación: {}'.format(estcontrato),
                                            cuerpo='Estimad@ {} se le informa que ha sido {} para la oferta {}'.format(
                                                contrato.persona.nombre_completo_minus(), estcontrato, contrato.oferta.titulo),
                                            destinatario_id=contrato.persona.id,
                                            # departamento=departamento,
                                            url='/emp_postulaciones',
                                            object_id=contrato.pk,
                                            prioridad=1,
                                            app_label='empleo',
                                            fecha_hora_visible=datetime.now() + timedelta(days=1)
                                            )
                notificacion.save(request)
                send_html_mail('Estado de contrato: {}/a'.format(estcontrato),
                               "emails/estadocontratopostulacionempleo.html",
                               {'sistema': u'Unemi Empleo', 'fecha': datetime.now(),
                                'contrato': contrato, 'tit': 'Unemi - Empleo', 'estcontrato': estcontrato},
                               contrato.persona.lista_emails_envio(),
                               [], cuenta=CUENTAS_CORREOS[17][1])
                log(u'Modifico estado de contratacion: %s' % contrato, request, "edit")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al contratar."})

        elif action == 'observacioncontrato':
            try:
                contrato = PersonaAplicaOferta.objects.get(pk=int(encrypt(request.POST['id'])))
                contrato.estcontrato = 2
                contrato.opc_contrato = request.POST['opc_contrato']
                if int(request.POST['opc_contrato']) == 0:
                    raise NameError('Debe seleccionar un motivo de rechazo')
                if int(request.POST['opc_contrato']) == 5:
                    contrato.observacioncontrato = request.POST['observacioncontrato']
                else:
                    contrato.observacioncontrato = OPCION_OBSERVACION_CONTRATO[int(request.POST['opc_contrato'])][1]
                contrato.fecha_contrato = hoy
                contrato.save()
                # estcontrato = 'No contratado'
                # send_html_mail('Estado de contrato: {}/a'.format(estcontrato),
                #                "emails/estadocontratopostulacionempleo.html",
                #                {'sistema': u'Unemi Empleo', 'fecha': datetime.now(),
                #                 'contrato': contrato, 'tit': 'Unemi - Empleo', 'estcontrato': estcontrato},
                #                contrato.persona.lista_emails_envio(),
                #                [], cuenta=CUENTAS_CORREOS[17][1])
                log(u'Modifico estado de contratacion: %s' % contrato, request, "edit")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al no contratar. {}".format(str(ex))})


        return JsonResponse({"result": False, "mensaje": u"Solicitud Incorrecta."})
    else:
        if 'action' in request.GET:
            data['action'] = action = request.GET['action']
            if action == 'add':
                try:
                    # puede_realizar_accion(request, 'empresa.puede_modificar_representantes')
                    data['title'] = u'Adicionar oferta laboral'
                    data['action'] = action
                    data['form'] = form = OfertaLaboralForm()
                    data['sbu'] = int(variable_valor('SBU_VALOR'))
                    form.adicionar(empresa)
                    return render(request, "empresa/addofertalaboral.html", data)
                except Exception as ex:
                    import sys
                    return JsonResponse({'resp': 'Error on line {} - {}'.format(sys.exc_info()[-1].tb_lineno, ex)})
            elif action == 'edit':
                try:
                    # puede_realizar_accion(request, 'sga.puede_modificar_administrativos')
                    data['title'] = u'Editar oferta laboral'
                    data['action'] = action
                    data['oferta'] = oferta = OfertaLaboralEmpresa.objects.get(pk=int(encrypt(request.GET['id'])))
                    carreras = []
                    for carrera in oferta.carrera.all():
                        item = {'id': carrera.id, 'text': capfirst(carrera.nombre.lower())}
                        carreras.append(item)
                    form = OfertaLaboralForm(initial=model_to_dict(oferta))
                    form.editar(oferta, empresa)
                    fecha_actual = datetime.now().date()
                    if fecha_actual > oferta.finicio:
                        form.bloquear_fechainicio()
                    data['id'] = int(encrypt(request.GET['id']))
                    data['carreras'] = carreras
                    data['form'] = form
                    data['sbu'] = int(variable_valor('SBU_VALOR'))
                    data['email_domain'] = EMAIL_DOMAIN
                    return render(request, "empresa/addofertalaboral.html", data)
                except Exception as ex:
                    pass
            elif action == 'buscarcarrera':
                try:
                    datos = []
                    id = request.GET['id']
                    carreras = Carrera.objects.filter(status=True, niveltitulacion_id=id, coordinacion__lte=5)
                    for carrera in carreras:
                        item = {'id': carrera.id, 'text': capfirst(carrera.nombre.lower())}
                        datos.append(item)
                    return JsonResponse({"result": "ok", 'lista': datos})
                except Exception as ex:
                    return JsonResponse({"result": "bad", 'mensaje': 'Error al obtener los datos'})
            elif action == 'buscartitulo':
                try:
                    datos = []
                    ids = json.loads(request.GET['ids'])
                    relacionadas = CarreraGruposCarrera.objects.filter(status=True, carrera__in=ids)
                    if relacionadas.exists():
                        for rel in relacionadas.first().carrera.all().values_list('id', flat=True):
                            ids.append(rel)
                    personascarrera = Graduado.objects.filter(status=True,inscripcion__status=True, inscripcion__carrera_id__in=ids, estadograduado=True).distinct('inscripcion__persona_id').values_list('inscripcion__persona_id', flat=True)
                    titulosids = Titulacion.objects.filter(status=True, persona_id__in=personascarrera, verificado=True, titulo__abreviatura__in=['ING','ING.', 'LIC','LIC.','T.S', 'C.P.A.', 'PS.', 'Q.F', 'ENF.', 'ING.LOG', 'MED.', 'TECN.', 'LCDO.', 'ECON.', 'LCA']).exclude(Q(titulo__nombre__icontains='EGRESAD')|Q(titulo__nombre__icontains='DIPLOMADO')).distinct().values_list('titulo_id', flat=True)
                    for titulo in Titulo.objects.filter(status=True, id__in=titulosids, nivel_id=3).distinct():
                        item = {'id': titulo.id, 'text': titulo.nombre}
                        datos.append(item)
                    return JsonResponse({"result": "ok", 'lista': datos})
                except Exception as ex:
                    return JsonResponse({"result": "bad", 'mensaje': 'Error al obtener los datos'})
            elif action == 'verdetalle':
                try:
                    data['id'] = id = int(encrypt(request.GET['id']))
                    data['filtro'] = filtro = OfertaLaboralEmpresa.objects.get(pk=id)
                    data['resp_campos'] = validar_campos(request, persona, filtro)
                    template = get_template("empresa/verdetalle.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass


            if action == 'observacioncontrato':
                 try:
                     data['id'] = id = int(encrypt(request.GET['id']))
                     data['filtro'] = filtro = PersonaAplicaOferta.objects.get(id=id)
                     data['form'] = GestionarContratoForm()
                     template = get_template("empresa/observacioncontrato.html")
                     return JsonResponse({"result": True, 'data': template.render(data)})
                 except Exception as ex:
                     pass



            if action == 'gestionarpostulacion':
                 try:
                     data['id'] = id = int(encrypt(request.GET['id']))
                     data['filtro'] = filtro = PersonaAplicaOferta.objects.get(id=id)
                     data['oferta'] = oferta = OfertaLaboralEmpresa.objects.get(id=int(encrypt(request.GET['ofer'])))
                     data['form'] = ObservacionHojaVidaForm()
                     template = get_template("empresa/observacionhojavida.html")
                     return JsonResponse({"result": True, 'data': template.render(data)})
                 except Exception as ex:
                     pass


            elif action == 'participantes':
                try:
                    data['title'] = 'Lista de postulantes'
                    data['id'] = id = int(encrypt(request.GET['id']))
                    data['oferta'] = oferta = OfertaLaboralEmpresa.objects.get(pk=id)
                    data['postulantes'] = postulantes = oferta.participantes()

                    estado, search, filtro, url_vars = request.GET.get('estado', ''), request.GET.get('search', ''), (Q(status=True)),f'&action={action}&id={request.GET["id"]}'

                    if estado:
                        data['estadoapto'] = int(estado)
                        url_vars += "&estado={}".format(estado)
                        postulantes = postulantes.filter(
                            Q(estado=estado))

                    if search:
                        data['search'] = search = request.GET['search'].strip()
                        url_vars += "&search={}".format(search)
                        postulantes = postulantes.filter(Q(persona__apellido1__icontains=search) | Q(persona__apellido2__icontains=search) | Q(persona__nombres__icontains=search) | Q(persona__cedula__icontains=search)).distinct()

                    paging = MiPaginador(postulantes, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['listado'] = page.object_list
                    data["url_vars"] = url_vars
                    data['list_count'] = len(postulantes)
                    data['estado'] = ESTADOS_APLICACION


                    #EXPORTAR EXCEL PARA VISUALIZAR POSTULANTES
                    if 'exportar_excel' in request.GET:
                        wb = openxl.Workbook()
                        wb["Sheet"].title = "Reporte_postulantes"
                        ws = wb.active
                        style_title = openxlFont(name='Arial', size=16, bold=True)
                        style_cab = openxlFont(name='Arial', size=10, bold=True)
                        alinear = alin(horizontal="center", vertical="center")
                        response = HttpResponse(content_type="application/ms-excel")
                        response[
                            'Content-Disposition'] = 'attachment; filename=Reporte de postulantes' + '-' + random.randint(
                            1, 10000).__str__() + '.xlsx'
                        ws.column_dimensions['B'].width = 35
                        ws.column_dimensions['C'].width = 15
                        ws.column_dimensions['D'].width = 15
                        ws.column_dimensions['E'].width = 25
                        ws.column_dimensions['F'].width = 25
                        ws.column_dimensions['G'].width = 25
                        ws.column_dimensions['H'].width = 20
                        ws.column_dimensions['I'].width = 20
                        ws.merge_cells('A1:I1')
                        ws['A1'] = 'LISTADO DE POSTULANTES'
                        celda1 = ws['A1']
                        celda1.font = style_title
                        celda1.alignment = alinear

                        ws.merge_cells('A2:B2')
                        ws['A2'] = 'Empresa: ' + oferta.empresa.nombre
                        celda1 = ws['A2']
                        celda1.font = style_cab
                        celda1.alignment = alinear

                        ws.merge_cells('A3:B3')
                        ws['A3'] = 'Oferta Laboral: ' + oferta.titulo
                        celda1 = ws['A3']
                        celda1.font = style_cab
                        celda1.alignment = alinear

                        columns = [u"N°", u"NOMBRE", u"CEDULA", u"TELEFONO", u"EMAIL",
                                   u"F. POSTULACIÓN", u"F. REVISIÓN", u"ESTADO", u"CONTRATADO",
                                   ]
                        row_num = 4
                        for col_num in range(0, len(columns)):
                            celda = ws.cell(row=row_num, column=(col_num + 1), value=columns[col_num])
                            celda.font = style_cab
                        row_num = 5
                        numero = 1
                        for list in postulantes:
                            ws.cell(row=row_num, column=1, value=numero)
                            ws.cell(row=row_num, column=2, value=str(list.persona.nombre_completo_minus()))
                            ws.cell(row=row_num, column=3, value=str(list.persona.cedula))
                            ws.cell(row=row_num, column=4, value=str(list.persona.telefono))
                            ws.cell(row=row_num, column=5, value=str(list.persona.emailinst) if list.persona.emailinst else list.persona.email)
                            ws.cell(row=row_num, column=6, value=str(list.fecha_creacion.date()))
                            ws.cell(row=row_num, column=7, value=str(list.fecha_revision.date()) if list.fecha_revision else 'Sin revisar')
                            ws.cell(row=row_num, column=8, value=str(list.get_estado_display()))
                            if list.estcontrato == 0:
                                estcontrato='Pendiente'
                            if list.estcontrato == 1:
                                estcontrato='Si'
                            if list.estcontrato == 2:
                                estcontrato='No'
                            ws.cell(row=row_num, column=9, value=str(estcontrato))
                            row_num += 1
                            numero += 1
                        wb.save(response)
                        return response


                    return render(request, "empresa/participanteslist.html", data)
                except Exception as ex:
                    pass
            elif action == 'verhojavida':
                try:
                    data['title'] = u'Hoja de vida del postulante'
                    hoy = datetime.now().date()
                    notas = None
                    if 'tab' in request.GET:
                        data['tab'] = request.GET['tab']
                    data['aspirante'] = aspirante = Persona.objects.get(id=int(encrypt(request.GET['id'])))
                    data['oferta'] = oferta = OfertaLaboralEmpresa.objects.get(id=int(encrypt(request.GET['ofer'])))
                    postulacion = PersonaAplicaOferta.objects.get(id=int(encrypt(request.GET['idpostu'])))
                    if postulacion.estado == 0:
                        postulacion.estado = 1
                        postulacion.fecha_revision = hoy
                        postulacion.save(request)
                    data['postulacion'] = postulacion
                    data['perfil'] = aspirante.mi_perfil()
                    # data['periodo'] = periodo
                    data['otrascertificaciones'] = CertificadoIdioma.objects.filter(status=True,
                                                                                    persona=aspirante).order_by('id')
                    data['niveltitulo'] = NivelTitulacion.objects.filter(pk__in=[3, 4], status=True).order_by('-rango')
                    data['desbloquear'] = desbloquear = False
                    data['v2'] = True
                    if oferta.muestrapromedio:
                        notas = Graduado.objects.filter(inscripcion__persona=aspirante)
                    data['notas'] = notas
                    return render(request, "empresa/verhojavida.html", data)
                except Exception as ex:
                    pass

            return HttpResponseRedirect(request.path)
        else:
            try:
                data['title'] = u'Ofertas Laborales'
                ofertas = OfertaLaboralEmpresa.objects.filter(status=True, empresa=empresa)
                data['fechaactual'] = datetime.now().date()
                search, filtro, url_vars, id = request.GET.get('search', ''), (Q(status=True)), '', request.GET.get('id', '')
                if search:
                    data['search'] = search = request.GET['search'].strip()
                    url_vars += "&search={}".format(search)
                    ofertas = ofertas.filter(Q(titulo__icontains=search) | Q(descripcion__icontains=search)).distinct()
                if id:
                    ofertas = ofertas.filter(id=int(encrypt(id)))
                    data['id'] = id
                paging = MiPaginador(ofertas.order_by('-finicio'), 20)
                p = 1
                try:
                    paginasesion = 1
                    if 'paginador' in request.session:
                        paginasesion = int(request.session['paginador'])
                    if 'page' in request.GET:
                        p = int(request.GET['page'])
                    else:
                        p = paginasesion
                    try:
                        page = paging.page(p)
                    except:
                        p = 1
                    page = paging.page(p)
                except:
                    page = paging.page(p)
                request.session['paginador'] = p
                data['paging'] = paging
                data['rangospaging'] = paging.rangos_paginado(p)
                data['page'] = page
                data['listado'] = page.object_list
                data["url_vars"] = url_vars
                data['list_count'] = len(ofertas)

                # EXPORTAR EXCEL PARA VISUALIZAR OFERTAS LABORALES
                if 'exportar_excel' in request.GET:
                    wb = openxl.Workbook()
                    wb["Sheet"].title = "Reporte_ofertas"
                    ws = wb.active
                    style_title = openxlFont(name='Arial', size=16, bold=True)
                    style_cab = openxlFont(name='Arial', size=10, bold=True)
                    alinear = alin(horizontal="center", vertical="center")
                    response = HttpResponse(content_type="application/ms-excel")
                    response[
                        'Content-Disposition'] = 'attachment; filename=Reporte de ofertas' + '-' + random.randint(
                        1, 10000).__str__() + '.xlsx'
                    ws.column_dimensions['B'].width = 40
                    ws.column_dimensions['C'].width = 25
                    ws.column_dimensions['D'].width = 25
                    ws.column_dimensions['E'].width = 20
                    ws.column_dimensions['F'].width = 30
                    ws.column_dimensions['G'].width = 25
                    ws.column_dimensions['H'].width = 10
                    ws.column_dimensions['I'].width = 22
                    ws.column_dimensions['J'].width = 23
                    ws.column_dimensions['K'].width = 20
                    ws.column_dimensions['L'].width = 20
                    ws.column_dimensions['M'].width = 20
                    ws.column_dimensions['N'].width = 20
                    ws.column_dimensions['O'].width = 20
                    ws.column_dimensions['P'].width = 20
                    ws.column_dimensions['Q'].width = 20

                    ws.merge_cells('A1:Q1')
                    ws['A1'] = 'LISTADO DE OFERTAS LABORALES'
                    celda1 = ws['A1']
                    celda1.font = style_title
                    celda1.alignment = alinear

                    ws.merge_cells('A2:B3')
                    ws['A2'] = 'Empresa: ' + empresa.nombre
                    celda1 = ws['A2']
                    celda1.font = style_cab
                    celda1.alignment = alinear

                    ws.merge_cells('E4:K4')
                    ws['E4'] = 'DETALLES DEL PUESTO'
                    celda2 = ws['E4']
                    celda2.font = style_cab
                    celda2.alignment = alinear

                    ws.merge_cells('L3:Q3')
                    ws['L3'] = 'CRONOGRAMA'
                    celda3 = ws['L3']
                    celda3.font = style_cab
                    celda3.alignment = alinear

                    ws.merge_cells('L4:M4')
                    ws['L4'] = 'PROCESO'
                    celda4 = ws['L4']
                    celda4.font = style_cab
                    celda4.alignment = alinear

                    ws.merge_cells('N4:O4')
                    ws['N4'] = 'POSTULACION'
                    celda5 = ws['N4']
                    celda5.font = style_cab
                    celda5.alignment = alinear

                    ws.merge_cells('P4:Q4')
                    ws['P4'] = 'REVISION'
                    celda6 = ws['P4']
                    celda6.font = style_cab
                    celda6.alignment = alinear

                    # ws.merge_cells('L4:M4')
                    # ws['L4'] = 'PROCESO'
                    # celda4 = ws['L4']
                    # celda4.font = style_cab
                    # celda4.alignment = alinear
                    #
                    # ws.merge_cells('N4:O4')
                    # ws['N4'] = 'POSTULACION'
                    # celda5 = ws['N4']
                    # celda5.font = style_cab
                    # celda5.alignment = alinear
                    #
                    # ws.merge_cells('P4:Q4')
                    # ws['P4'] = 'REVISION'
                    # celda6 = ws['P4']
                    # celda6.font = style_cab
                    # celda6.alignment = alinear

                    columns = [u"N°", u"NOMBRE", u"ENCARGADO", u"FORMACION ACADEMICA", u"MODALIDAD",
                               u"DEDICACION", u"JORNADA", u"RMU", u"TIPO DE CONTRATO", u"EXPERIENCIA", u"VACANTES",
                               u"INICIO", u"FIN", u"INICIO", u"FIN", u"INICIO", u"FIN"
                               ]
                    row_num = 5
                    for col_num in range(0, len(columns)):
                        celda = ws.cell(row=row_num, column=(col_num + 1), value=columns[col_num])
                        celda.font = style_cab
                    row_num = 6
                    numero = 1
                    for list in ofertas:
                        ws.cell(row=row_num, column=1, value=numero)
                        ws.cell(row=row_num, column=2, value=str(list.titulo))
                        ws.cell(row=row_num, column=3, value=str(list.encargado.persona.nombre_completo_minus()))
                        ws.cell(row=row_num, column=4, value=str(NIVEL_INSTRUCCION[0][1]) if list.nivel == 3 else str(NIVEL_INSTRUCCION[1][1]))
                        ws.cell(row=row_num, column=5, value=str(MODALIDAD[list.modalidad][1]))
                        ws.cell(row=row_num, column=6, value=str(DEDICACION[list.dedicacion][1]))
                        ws.cell(row=row_num, column=7, value=str(JORNADA[list.jornada][1]))
                        ws.cell(row=row_num, column=8, value=str(list.rmu))
                        ws.cell(row=row_num, column=9, value=str(list.tipocontrato.nombre))
                        ws.cell(row=row_num, column=10, value=str(TIEMPO[list.tiempoexperiencia][1]))
                        ws.cell(row=row_num, column=11, value=str(list.vacantes))
                        ws.cell(row=row_num, column=12, value=str(list.finicio)if list.finicio else 'Sin asignar')
                        ws.cell(row=row_num, column=13,
                                value=str(list.ffin)if list.ffin else 'Sin asignar')
                        ws.cell(row=row_num, column=14, value=str(list.finiciopostulacion)if list.finiciopostulacion else 'Sin asignar')
                        ws.cell(row=row_num, column=15, value=str(list.ffinpostlacion)if list.ffinpostlacion else 'Sin asignar')
                        ws.cell(row=row_num, column=16, value=str(list.finiciorevision)if list.finiciorevision else 'Sin asignar')
                        ws.cell(row=row_num, column=17, value=str(list.ffinrevision)if list.ffinrevision else 'Sin asignar')
                        # ws.cell(row=row_num, column=5,
                        #         value=str(list.persona.emailinst) if list.persona.emailinst else list.persona.email)
                        # ws.cell(row=row_num, column=6, value=str(list.fecha_creacion.date()))
                        # ws.cell(row=row_num, column=7,
                        #         value=str(list.fecha_revision) if list.fecha_revision else 'Sin revisar')
                        # ws.cell(row=row_num, column=8, value=str(list.get_estado_display()))
                        row_num += 1
                        numero += 1
                    wb.save(response)
                    return response

                return render(request, "empresa/viewOfertaLaboral.html", data)
            except Exception as ex:
                pass



# def enviar_notificaciones_postulantes(oferta):
#     try:
#         tipo = oferta.quienpostula
#
#         filtro = Q(status=True)
#         if tipo == 0:
#             filtro += filtro & Q(graduado__status=True, graduado__estadograduado=True)
#         elif tipo == 1:
#             filtro += filtro & Q(inscripcionnivel__nivel_id__gte=7)
#         else:
#             filtro += filtro & Q(graduado__status=True, graduado__estadograduado=True)
#             filtro += filtro & Q(inscripcionnivel__nivel_id__gte=7)
#         if oferta.carrera.exists():
#             if oferta.carrerarelacionada.exists():
#                 filtro += filtro & Q(Q(carrera_id__in=oferta.carrera.values_list('id', flat=True)) | Q(carrera_id__in=oferta.carrerarelacionada.values_list('id', flat=True)))
#             else:
#                 filtro += filtro & Q(carrera_id__in=oferta.carrera.values_list('id', flat=True))
#         if oferta.sexo > 0:
#             filtro += filtro & Q(persona__seso_id=oferta.sexo)
#         if oferta.discapacitados:
#             filtro += filtro & Q(perfilinscripcion__tienediscapacidad=True, perfilinscripcion__status=True, perfilinscripcion__verificadiscapacidad=True)
#
#         notificacion(asunto, postar.observacion, para, None,
#                      '/alu_solicitudcambiocarrera?action=verproceso&id={}'.format(
#                          encrypt(postar.solicitud.pk)), postar.pk, 1, 'sga',
#                      DocumentosSolicitudCambioCarrera, request)
#
#     except Exception as e:
#         print(e)