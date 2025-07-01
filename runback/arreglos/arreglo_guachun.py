import os
import statistics
import sys
import django

# SITE_ROOT = os.path.dirname(os.path.realpath(__file__))
from unidecode import unidecode

# Configuración de Django
YOUR_PATH = os.path.dirname(os.path.realpath(__file__))
# print(f"YOUR_PATH: {YOUR_PATH}")
SITE_ROOT = os.path.dirname(os.path.dirname(YOUR_PATH))
SITE_ROOT = os.path.join(SITE_ROOT, '')
# print(f"SITE_ROOT: {SITE_ROOT}")
your_djangoproject_home = os.path.split(SITE_ROOT)[0]
# print(f"your_djangoproject_home: {your_djangoproject_home}")
sys.path.append(your_djangoproject_home)

from webpush import send_user_notification
from django.core.wsgi import get_wsgi_application

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "settings")
application = get_wsgi_application()
import calendar
import datetime
from django.db import transaction
from sga.models import *
from sagest.models import *
from django.db.models import Sum, F, FloatField, IntegerField
from django.db.models.functions import Coalesce, ExtractYear
from settings import MEDIA_ROOT, BASE_DIR
from gdocumental.models import *
from bd.models import *
from balcon.models import EncuestaProceso, RespuestaEncuestaSatisfaccion
from empleo.models import ResponsableConvenio
# Trato de documentos xls
# Pandas libreria reciente más legible
import pandas as pd

# Openpyxl libreria para tratar cantidades grandes de registros e integración sencilla de graficos
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, colors
django.setup()

from sga.models import RecordAcademico, Malla, MateriaAsignada

def actualizar_activos_fijos():
    with transaction.atomic():
        try:
            archivo_ = 'activos_fijos_actualizar'
            url_archivo = "{}/media/{}.xlsx".format(SITE_STORAGE, archivo_)
            wb = openpyxl.load_workbook(filename=url_archivo)
            # ws = wb.get_sheet_by_name(archivo[1]) # Permite seleccionar una hoja en especifico del excel
            ws = wb.active  # Selecciona la hoja activa u hoja principal
            total_filas = ws.max_row - 2
            for row in ws.iter_rows(min_row=3):
                c_gobierno = row[0].value
                c_interno = int(row[1].value) if row[1].value and int(row[1].value) != 0 else ''
                filtro = Q(status=True)
                if c_gobierno:
                    filtro = filtro & Q(codigogobierno=c_gobierno)
                if c_interno:
                    filtro = filtro & Q(codigointerno=c_interno)
                if c_gobierno or c_interno:
                    if ActivoFijo.objects.filter(filtro).exists():
                        activo_f = ActivoFijo.objects.get(filtro)
                        activo_f.vidautil = row[3].value
                        activo_f.costo = row[5].value
                        activo_f.valorresidual = row[6].value
                        activo_f.valorlibros = row[7].value
                        activo_f.valordepreciacionacumulada = row[8].value
                        activo_f.save()
                        print(f'Código: {row[0].value}, '
                              f'Vida útil: {row[3].value}, '
                              f'Valor contable: {row[5].value}, '
                              f'Valor residual: {row[6].value}, '
                              f'Valor en libros: {row[7].value}, '
                              f'Valor depreciación acumulada: {row[8].value}')
                    else:
                        raise NameError(f'Error con c_gobierno: {c_gobierno}, c_interno:{c_interno}')
                else:
                    total_filas -= 1
            print(f'Se actulizo: {total_filas}')
        except Exception as ex:
            transaction.set_rollback(True)
            print(str(ex))
# actualizar_activos_fijos()

def actualizar_estado_edcon():
    inscritos = CapInscritoIpec.objects.filter(status=True)
    cont = 0
    for i in inscritos:
        nota_final = i.instructor_notasfinales()
        if nota_final[0][1] == 'APROBADO':
            cont += 1
            i.estado = 2
            i.save()
            print(f'{i.id} Estado cambiado: aprobado')
    print(cont)
# actualizar_estado_edcon()

def actualizar_calculo_valores_producto():
    with transaction.atomic():
        try:
            hoy = datetime.now()
            anio_anterior = hoy.date().year - 1
            inicio, fin = datetime(anio_anterior, 1, 1), datetime(anio_anterior, 12, 31)
            tiempo_reposicion = 15
            total_actualizado = 0
            productos_ids = KardexInventario.objects.filter(fecha__range=(inicio, fin), tipomovimiento=2, status=True).values_list('producto_id').distinct()
            productos = Producto.objects.filter(status=True).exclude(id__in=productos_ids)
            for p in productos:
                valores_salida = []
                total_salida = 0
                # CÁLCULO DE VALORES SALIDA
                ultimo_kardex = KardexInventario.objects.annotate(anio=ExtractYear('fecha')).filter(anio__lt=hoy.date().year, tipomovimiento=2, producto_id=p.id, status=True).order_by('fecha').last()
                if not ultimo_kardex:
                    ultimo_kardex = KardexInventario.objects.annotate(anio=ExtractYear('fecha')).filter(tipomovimiento=2, producto_id=p.id, status=True).order_by('fecha').last()
                if ultimo_kardex:
                    anio = ultimo_kardex.anio
                    ini, fi = datetime(anio, 1, 1), datetime(anio, 12, 31)
                    kardexs = KardexInventario.objects.filter(fecha__range=(ini, fi), tipomovimiento=2, producto_id=p.id, status=True)
                    for mes in range(1, 13):
                        ultimo_dia = calendar.monthrange(anio, mes)[1]
                        i_mes = datetime(anio, mes, 1)
                        f_mes = datetime(anio, mes, ultimo_dia)
                        kardexs_mes = kardexs.filter(fecha__range=(i_mes, f_mes)).aggregate(valor_salida=Sum('valor'), total_cantidad=Sum('cantidad'))
                        valor_salida = round(kardexs_mes['valor_salida'], 2) if kardexs_mes['valor_salida'] else 0
                        total_cantidad = round(kardexs_mes['total_cantidad'], 2) if kardexs_mes['total_cantidad'] else 0
                        total_salida += valor_salida
                        diccionario = {'mes': mes, 'valor_salida': valor_salida, 'total_cantidad': total_cantidad}
                        valores_salida.append(diccionario)
                    consumo_minimo = round(min(valores_salida, key=lambda x: x["valor_salida"])["valor_salida"], 2)
                    consumo_maximo = round(max(valores_salida, key=lambda x: x["valor_salida"])["valor_salida"], 2)
                    consumo_medio = round(statistics.mean([float(d["valor_salida"]) for d in valores_salida]), 2)
                    consumo_minimo = consumo_minimo if not consumo_minimo == 0 else 1
                    cantidad_minima = min(valores_salida, key=lambda x: x["total_cantidad"])["total_cantidad"]
                    kardex_maxima = max(valores_salida, key=lambda x: x["total_cantidad"])["total_cantidad"]
                    kardex_minima = cantidad_minima if not cantidad_minima == 0 else 1

                    p.consumo_minimo_diario = consumo_minimo
                    p.consumo_medio_diario = consumo_medio
                    p.consumo_maximo_diario = consumo_maximo
                    p.kardex_maximo = kardex_maxima
                    p.kardex_minimo = kardex_minima
                    p.tiempo_reposicion_inventario = tiempo_reposicion
                    p.save()
                    p.minimo = p.calcular_existencia_minima()
                    p.maximo = p.calcular_existencia_maxima()
                    p.save()
                    total_actualizado += 1
                    totales = {'id_producto': p.id, 'total_salida': total_salida,
                               'consumo_minimo': consumo_minimo,
                               'consumo_maximo': consumo_maximo,
                               'consumo_medio': consumo_medio,
                               'cantidad_minima': kardex_minima,
                               'cantidad_maxima': kardex_maxima,
                               }
                    print(totales)
            print(f'Actualizado:{total_actualizado}')
            print(f'Sin actualizar:{len(productos) - total_actualizado}')
        except Exception as ex:
            transaction.set_rollback(True)
            print(str(ex))
# actualizar_calculo_valores_producto()

def secuencia_activos():
    if not SecuenciaActivos.objects.exists():
        secuencia = SecuenciaActivos()
        secuencia.save()
        return secuencia
    else:
        return SecuenciaActivos.objects.all()[0]

def crear_constataciones_mal_estado():
    with transaction.atomic():
        try:
            activosfijo = ActivoFijo.objects.filter(status=True, procesobaja=True)
            creados, actualizados = 0, 0
            for activo in activosfijo:
                detalle_c = DetalleConstatacionFisica.objects.filter(status=True, activo=activo, codigoconstatacion__periodo_id=1).first()
                if not detalle_c:
                    constatacion = ConstatacionFisica.objects.filter(status=True, usuariobienes=activo.responsable, periodo_id=1).first()
                    if not constatacion:
                        secuencia = secuencia_activos()
                        secuencia.numeroconstatacion += 1
                        secuencia.save()
                        constatacion = ConstatacionFisica(usuariobienes=activo.responsable,
                                                          numero=secuencia.numeroconstatacion,
                                                          normativaconstatacion=secuencia.normativaconstatacion,
                                                          fechainicio=datetime.now(),
                                                          periodo_id=1,
                                                          ubicacionbienes=activo.ubicacion)
                        constatacion.save()
                    detalle_c = DetalleConstatacionFisica(codigoconstatacion=constatacion,
                                                          activo=activo,
                                                          responsable_id=1204,
                                                          ubicacionbienes=activo.ubicacion,
                                                          estadooriginal=activo.estado,
                                                          estadoactual_id=3,
                                                          enuso=False,
                                                          encontrado=True)
                    detalle_c.save()
                    creados += 1
                    print(f'Constatación creada: {creados}')
                else:
                    detalle_c.ubicacionbienes = activo.ubicacion
                    detalle_c.estadooriginal = activo.estado
                    detalle_c.estadoactual_id = 3
                    detalle_c.enuso = False
                    detalle_c.encontrado = True
                    detalle_c.save()
                    actualizados += 1
                    print(f'Constatación actualizada: {actualizados}')
            print(f'Se constatado :{len(activosfijo)} activos, Creados: {creados}, Actualizados:{actualizados}')
        except Exception as ex:
            transaction.set_rollback(True)
            print(str(ex))
# crear_constataciones_mal_estado()

def actualizar_numero_hijos():
    try:
        personas = DistributivoPersona.objects.filter(status=True)
        for d in personas:
            total_hijos = len(d.persona.cargas())
            per_extension = d.persona.personaextension_set.filter(status=True).first()
            per_extension.hijos = total_hijos
            per_extension.save(update_fields=["hijos"])
            print(f'Cantidad de hijos actualizado: {total_hijos}')
        print(f'Cantidad personas actualizadas: {len(personas)}')
    except Exception as ex:
        transaction.set_rollback(True)
        print(str(ex))
# actualizar_numero_hijos()

def migrar_estado():
    try:
        print(f'Inicio proceso')
        cronogramas = CronogramaPersonaConstatacionAT.objects.filter(status=True, periodo_id=2, estado=4)
        cont=0
        for cronograma in cronogramas:
            activos_c = cronograma.detalleconstatacionfisicaactivotecnologico_set.filter(status=True, cronograma__status=True, constatado=True)
            for ac in activos_c:
                activo = ac.activo
                if not activo.estado == ac.estadoactual:
                    cont += 1
                    activo.estado = ac.estadoactual
                    activo.save(update_fields=["estado"])
                    print(f'{cont}. Estado actualizado codigogobierno={activo.activotecnologico.codigogobierno} activo_id={activo.id} - {activo.estado} de {cronograma.persona}')
        print(f'Se actualizo satisfactoriamente {cont} de {len(cronogramas)} personas')
    except Exception as ex:
        transaction.set_rollback(True)
        print(str(ex))
# migrar_estado()

def actualizar_encuesta_balcon():
    try:
        print('Inicio proceso de actualización de content type y object_id')
        e_procesos = EncuestaProceso.objects.filter(proceso__isnull=False, status=True)
        cont = 0
        for ep in e_procesos:
            content_type = ContentType.objects.get_for_model(ep.proceso)
            ep.content_type = content_type
            ep.categoria_id = 2
            ep.object_id = ep.proceso.id
            ep.save()
            cont += 1
            respuestas = RespuestaEncuestaSatisfaccion.objects.filter(pregunta__encuesta=ep, status=True, solicitud__isnull=False)
            for respuesta in respuestas:
                content_type = ContentType.objects.get_for_model(respuesta.solicitud)
                respuesta.content_type = content_type
                respuesta.object_id = respuesta.solicitud.id
                respuesta.save()
            print(f'{cont}. Modelo: {content_type}, id: {ep.id} - añadido y {len(respuestas)} respuestas actualizadas')
        print(f'{len(e_procesos)} actualizados.')
    except Exception as ex:
     transaction.set_rollback(True)
     print(str(ex))
# actualizar_encuesta_balcon()

def extraer_valores_excel():
    try:
        archivo_ = 'BATERIAS_BAJA_2019'
        url_archivo = "D:\\git\\academico\\media\\BATERIAS_BAJA_2019.xlsx"
        nombres_hojas = pd.ExcelFile(url_archivo).sheet_names
        for name in nombres_hojas:
            url_guardar = f"D:\\git\\academico\\media\\{name}.xlsx"
            df = pd.read_excel(url_archivo, sheet_name=name)
            # if not 'CÓDIGO DEL BIEN' in df.columns:
            #     raise NameError('Formato de archivo erróneo, columna código del bien faltante.')
            df['Valor contable'] = ""
            df['Fecha ingreso'] = ""
            df['Fecha baja'] = ""
            df['Códigos gob iguales'] = ""
            df['Códigos int iguales'] = ""
            cont=0
            for index, row in df.iterrows():
                codigo = str(row['CÓDIGO']).strip().split('.')[0]
                codigo_interno=''
                activo = ActivoFijo.objects.filter(Q(codigogobierno__iexact=codigo) |
                                                   Q(codigointerno__iexact=codigo)).first()
                if 'CÓDIGO ANTERIOR' in df.columns:
                    codigo_interno = str(row['CÓDIGO ANTERIOR']).strip().split('.')[0]
            print(f'Activos no encontrados: {cont}, encontrados:')
            df.to_excel(url_guardar, index=False)
    except Exception as ex:
        print(f'{ex}')

# extraer_valores_excel()

def extraer_valores_contables_excel():
    try:
        archivo_ = 'adjunto'
        url_archivo = "D:\\git\\academico\\media\\adjunto.xlsx"
        url_guardar = f"D:\\git\\academico\\media\\acta_entrega.xlsx"
        df = pd.read_excel(url_archivo, sheet_name='ACTA DE ENTREGA')
        # if not 'CÓDIGO DEL BIEN' in df.columns:
        #     raise NameError('Formato de archivo erróneo, columna código del bien faltante.')
        df['Valor contable'] = ""
        df['Fecha ingreso'] = ""
        cont=0
        for index, row in df.iterrows():
            codigo = str(row['Cod. gobierno']).strip().split('.')[0]
            codigo_interno = str(row['Cod. interno']).strip().split('.')[0]
            activo = None
            if codigo:
                activo = ActivoFijo.objects.filter(Q(codigogobierno__iexact=codigo) |
                                                   Q(codigointerno__iexact=codigo)).first()

            if codigo_interno and not activo:
                activo = ActivoFijo.objects.filter(Q(codigointerno__iexact=codigo_interno) |
                                                   Q(codigogobierno__iexact=codigo_interno)).first()

            if not activo:
                cont += 1
                print(f'Activo no encontrado {codigo}')
            else:
                baja = DetalleBajaActivo.objects.filter(status=True, activo=activo, seleccionado=True).first()
                df.at[index, 'Valor contable'] = f'$ {activo.costo}'
                df.at[index, 'Fecha ingreso'] = activo.fechaingreso
                print(f'Información extraida correctamente {codigo}')
        print(f'Activos no encontrados: {cont}, encontrados:')
        df.to_excel(url_guardar, index=False)
    except Exception as ex:
        print(f'{ex}')
# extraer_valores_contables_excel()

def migrar_responsables_internos_convenios():
    try:
        cont=0
        print('Inicio proceso de migración')
        convenios = ConvenioEmpresa.objects.filter(status=True, responsableinterno__isnull=False)
        for convenio in convenios:
            cargo = convenio.cargo_denominaciones.all()
            cargo = cargo.first() if len(cargo) == 1 else None
            if not cargo:
                cargo = convenio.responsableinterno.mi_cargo_administrativo()
                if not cargo:
                    cargo = convenio.responsableinterno.distributivopersonahistorial_set.filter(status=True, regimenlaboral_id=1).order_by('-fecha_creacion').first()
                    cargo = cargo.denominacionpuesto if cargo else convenio.responsableinterno.mi_cargo()
            if not ResponsableConvenio.objects.filter(convenio=convenio, persona=convenio.responsableinterno):
                responsable = ResponsableConvenio(convenio=convenio, persona=convenio.responsableinterno, cargo=cargo)
                responsable.save()
                cont += 1
                print(f'Se migro el responsable con exito: {responsable}')
        convenios = ConvenioEmpresa.objects.filter(status=True, responsableinterno__isnull=True)
        print(f'{cont} convenios de responsable interno Migrados.')
        cont = 0
        for convenio in convenios:
            distributivos = DistributivoPersona.objects.filter(denominacionpuesto__in=convenio.cargo_denominaciones.all().values_list('id', flat=True), status=True)
            for distributivo in distributivos:
                if not ResponsableConvenio.objects.filter(convenio=convenio, persona=distributivo.persona):
                    responsable = ResponsableConvenio(convenio=convenio, persona=distributivo.persona, cargo=distributivo.denominacionpuesto)
                    responsable.save()
                    cont += 1
                    print(f'Se migro el responsable con exito: {responsable}')
        print(f'{cont} convenios por cargo Migrados.')
    except Exception as ex:
        transaction.set_rollback(True)
        print(str(ex))

# migrar_responsables_internos_convenios()

def crear_sesiones_resolucion():
    try:
        cont=0
        print('Inicio proceso de migración')
        resoluciones = Resoluciones.objects.filter(status=True)
        fechas = resoluciones.values_list('fecha', 'tipo_id').order_by('fecha', 'tipo_id').distinct()
        for r in fechas:
            cont += 1
            fecha = r[0].strftime("%Y-%m-%d")
            nombre = f'Sesión: {fecha}'
            sesion = SesionResolucion(tipo_id=int(r[1]), fecha=fecha, nombre=nombre)
            sesion.save()
            resoluciones = Resoluciones.objects.filter(status=True, fecha=r[0], tipo=sesion.tipo)
            for resolucion in resoluciones:
                resolucion.sesion=sesion
                resolucion.save()
            print(f'Sesión creada correctamente {cont}')
        transaction.rollback()
        print(f"Ocurrió un error: {str(ex)}")

        print(f'{len(fechas)} Sesiones creadas')
    except Exception as ex:
        transaction.set_rollback(True)
        print(str(ex))

#crear_sesiones_resolucion()

def mover_constataciones_duplicaddas():
    try:
        cont=0
        print('Inicio proceso de constatacion')
        responsables = ConstatacionFisica.objects.filter(periodo_id=1, estado=1).values_list('usuariobienes_id', flat=True).order_by('usuariobienes_id').distinct()
        for responsable in responsables:
            constataciones = ConstatacionFisica.objects.filter(usuariobienes=responsable, periodo_id=1, estado=1).order_by('numero')
            if len(constataciones) > 1:
                first_constatacion = constataciones.first()
                constataciones = constataciones.exclude(id=first_constatacion.id)
                for constatacion in constataciones:
                    constatacion.detalle_constatacion().update(codigoconstatacion=first_constatacion)
                    # constatacion.delete()
                    print(f'Cabecera de constatación eliminara y sus detalles actualizados a primera cabcera salvada {first_constatacion}')

        constataciones = ConstatacionFisica.objects.filter(status=True, periodo__isnull=True, estado=2).order_by('numero')
        numero = constataciones.last().numero
        constataciones = ConstatacionFisica.objects.filter(periodo_id=1, estado=1).order_by('numero')
        for constatacion in constataciones:
            numero += 1
            if not constatacion.numero == numero:
                constatacion.numero = numero
                constatacion.save()
                print(f'Constatación actualizada {numero} - {constatacion}')
        print(f'Finalizado con éxito')
    except Exception as ex:
        transaction.set_rollback(True)
        print(str(ex))

# mover_constataciones_duplicaddas()

def delete_cabeceraconstatacion():
    try:
        print('Inicio proceso de constatacion')
        constataciones = ConstatacionFisica.objects.filter(periodo_id=1, estado=1).order_by('numero')
        cont = 0
        for constatacion in constataciones:
            if not constatacion.detalle_constatacion().exists():
                cont += 1
                constatacion.delete()
                print(f'{cont}.Cabecera eliminada')
        constataciones = ConstatacionFisica.objects.filter(status=True, periodo__isnull=True, estado=2).order_by('numero')
        numero = constataciones.last().numero
        constataciones = ConstatacionFisica.objects.filter(periodo_id=1, estado=1).order_by('numero')
        for constatacion in constataciones:
            numero += 1
            if not constatacion.numero == numero:
                constatacion.numero = numero
                constatacion.save()
                print(f'Constatación actualizada {numero} - {constatacion}')
        print(f'Finalizado con éxito')
    except Exception as ex:
        transaction.set_rollback(True)
        print(str(ex))

# delete_cabeceraconstatacion()

def ordenar_sessiones_resoluciones():
    try:
        print('Inicio proceso de ordenamiento')
        tipos = TipoResolucion.objects.filter(status=True)
        for t in tipos:
            sesiones = t.sesiones().order_by('fecha_creacion')
            for s in sesiones:
                if s.orden == 0:
                    s.orden = s.orden_next()
                    s.save()
                    print(f'Sesión {s}--{s.orden}')
                for idx, r in enumerate(s.resoluciones()):
                    orden = r.numeroresolucion[-3:]
                    i = -1  # Índice para recorrer hacia la izquierda
                    # Encontrar la posición del primer carácter no numérico
                    while i >= -len(orden) and orden[i].isdigit():
                        i -= 1
                    # Extraer los dígitos numéricos y convertir a entero
                    num_part = orden[i + 1:]
                    if num_part.isdigit():
                        r.orden = int(num_part)
                    else:
                        r.orden = idx + 1
                    r.save()
                    print(f'Resolución {r.numeroresolucion}--{r.orden}')
            resoluciones = Resoluciones.objects.filter(status=True, sesion__isnull=True, tipo=t).order_by('fecha_creacion')
            for idx, r in enumerate(resoluciones):
                orden = r.numeroresolucion[-3:]
                i = -1  # Índice para recorrer hacia la izquierda
                # Encontrar la posición del primer carácter no numérico
                while i >= -len(orden) and orden[i].isdigit():
                    i -= 1
                # Extraer los dígitos numéricos y convertir a entero
                num_part = orden[i + 1:]
                if num_part.isdigit():
                    r.orden = int(num_part)
                else:
                    r.orden = idx
                r.save()
                print(f'Resolución {r.numeroresolucion}--{r.orden}')
        print(f'Proceso finalizado')
    except Exception as ex:
        transaction.set_rollback(True)
        print(str(ex))

# ordenar_sessiones_resoluciones()

# Lógica de modelo para Preguntas de Paz Y Salvo
# Modelo DetalleDireccionFormatoPS, atributo logicamodelo
def verificar_pregunta(pazsalvo):
    from sagest.models import ActivoFijo
    funcionario = pazsalvo.persona
    activos = ActivoFijo.objects.filter(Q(responsable=funcionario) |
                                        Q(custodio=funcionario),
                                        status=True).exists()
    cargos = funcionario.mis_cargos().exclude(denominacionpuesto=pazsalvo.cargo,
                                              fecha_creacion__lte=pazsalvo.fecha).exists()
    return not activos or cargos

## Lógica de modelo para Preguntas de Paz Y Salvo De operaciones
# Nombre de la funcion tiene que ser verificar_pregunta
def verificar_pregunta_at(pazsalvo):
    from sagest.models import ActivoFijo
    funcionario = pazsalvo.persona
    activos = ActivoFijo.objects.filter(responsable=funcionario, status=True).exists()
    cargos = funcionario.mis_cargos().exclude(denominacionpuesto=pazsalvo.cargo,
                                              fecha_creacion__lte=pazsalvo.fecha).exists()
    return not activos or cargos


def crear_respuestas_jefe_ps():
    from sagest.models import PazSalvo, DetallePazSalvo
    try:
        print('Inicio de proceso')
        pazsalvos = PazSalvo.objects.filter(status=True, estado_requisito=1)
        cont=0
        for idx, ps in enumerate(pazsalvos):
            jefe = ps.jefeinmediato
            cargo = jefe.mi_cargo_administrativo() if jefe.mi_cargo_administrativo() else jefe.mi_cargo()
            cumplimiento = ps.cumplimiento(cargo.id, True)
            if not cumplimiento['respondio']:
                preguntas = ps.preguntas(cargo.id, True, True)
                preguntas_all = ps.preguntas_all()
                respuestas_all = ps.respuestas_all()
                total = len(preguntas_all) - len(respuestas_all)
                if total > len(preguntas):

                    for idxre, pregunta in enumerate(preguntas):
                        respuesta = ps.respuestas_jefe().filter(pregunta=pregunta).first()
                        if not respuesta:
                            respuesta = DetallePazSalvo(persona=jefe,
                                                        pazsalvo=ps,
                                                        pregunta=pregunta,
                                                        marcado=True)
                            respuesta.save()
                            print(f'{idx}.{idxre}. Se crea respuestas de jefe inmediato Paz y Salvo (pazsalvo_id : {ps.id} | respuesta_id: {respuesta.id})')
                            cont += 1
        print(f'Finalizo proceso de creación con éxito, se creo {cont}')
    except Exception as ex:
        transaction.set_rollback(True)
        print(f'{ex}')

# crear_respuestas_jefe_ps()

def crear_respuestas_financiero_ps():
    from sagest.models import PazSalvo, DetallePazSalvo
    try:
        print('Inicio de proceso')
        pazsalvos = PazSalvo.objects.filter(status=True, estado=2)
        cont=0
        for idx, ps in enumerate(pazsalvos):
            jefe = Persona.objects.get(id=8294)
            cargo = jefe.mi_cargo_administrativo() if jefe.mi_cargo_administrativo() else jefe.mi_cargo()
            cumplimiento = ps.cumplimiento(cargo.id)
            if not cumplimiento['respondio']:
                preguntas = ps.preguntas(cargo.id)
                for idxre, pregunta in enumerate(preguntas):
                    respuesta = ps.respuestas(cargo).filter(pregunta=pregunta, persona=jefe).first()
                    if not respuesta:
                        respuesta = DetallePazSalvo(persona=jefe,
                                                    pazsalvo=ps,
                                                    pregunta=pregunta,
                                                    respondio=True,
                                                    marcado=True)
                        respuesta.save()
                        cont += 1
        print(f'Finalizo proceso de creación con éxito, se creo {cont}')
    except Exception as ex:
        transaction.set_rollback(True)
        print(f'{ex}')

# crear_respuestas_financiero_ps()

def actualizar_estado_actas_postulate():
    from postulate.models import ActaPartida
    try:
        print('Inicio de proceso')
        actas = ActaPartida.objects.filter(status=True).exclude(estado=3)
        for idx, acta in enumerate(actas):
            ids_responsables = acta.responsables().values_list('id', flat=True).order_by('id').distinct()
            firmados = acta.historial_firmados().filter(personatribunal_id__in=ids_responsables).values_list('personatribunal_id', flat=True).order_by('personatribunal_id').distinct('personatribunal_id')
            acta.estado = 3 if len(firmados) >= len(ids_responsables) else 2
            acta.save()
        print(f'Finalizo proceso de creación con éxito')
    except Exception as ex:
        transaction.set_rollback(True)
        print(f'{ex}')

def revertir_estado_ps():
    from sagest.models import PazSalvo
    try:
        print('Inicio de proceso')
        pazsalvos = PazSalvo.objects.filter(status=True, estado=2)
        for idx, ps in enumerate(pazsalvos):
            respuestas_all = len(ps.detallepazsalvo_set.filter(status=True, respondio=True).order_by('pregunta_id').distinct('pregunta_id'))
            if not len(ps.preguntas_all()) <= respuestas_all:
                ps.estado = 1
                ps.save()
                print(f'{ps.id}. Actualizado')
        print(f'Finalizo proceso de creación con éxito')
    except Exception as ex:
        transaction.set_rollback(True)
        print(f'{ex}')

# revertir_estado_ps()
def cambiar_estado():
    from sagest.funciones import encrypt_id
    constataciones= ConstatacionFisica.objects.filter(id=encrypt_id('OPPQQRRSSTTUUVVWRVUP'))
    cont=0
    for constatacion in constataciones:
        cont+=1
        acta = ActaConstatacion.objects.filter(status=True,constatacion=constatacion)
        acta.update(status=False)
        constatacion.estadoacta=1
        constatacion.estado=1
        constatacion.save()

# cambiar_estado()

def actualizar_accionpersonal():
    try:
        # ACTUALIZAR UNIDADES
        archivo_unidades = 'unidad_organica.xlsx'
        path_file = os.path.join(SITE_ROOT, 'media', archivo_unidades)
        name_hoja = pd.ExcelFile(path_file).sheet_names[0]
        df = pd.read_excel(path_file, sheet_name=name_hoja)
        cont_u = 0
        for index, row in df.iterrows():
            unidad_old = unidecode(str(row['unidad_old']).strip().replace('  ', ' '))
            unidad_now = unidecode(str(row['unidad_now']).strip().replace('  ', ' '))
            ePersonaAcciones = PersonaContratos.objects.filter(status=True, unidadorganica__isnull=True, unidad__unaccent__iexact=unidad_old)
            for contrato in ePersonaAcciones:
                departamento = Departamento.objects.filter(status=True, nombre__unaccent__iexact=unidad_now).first()
                if not departamento:
                    departamento = Departamento(nombre=unidad_now)
                    departamento.save()
                    # print(f'Se creo nuevo departamento, {unidad_now}')
                    print(f'Se creo nuevo departamento {departamento.id} | {unidad_now}')
                # contrato.unidad = unidad_now
                contrato.unidadorganica = departamento
                contrato.save(update_fields=['unidad', 'unidadorganica'])
                cont_u += 1
                print(f'{contrato.id}) Se edito unidad de {contrato}-{contrato.persona} | Anterior={unidad_old} <----> Actual={unidad_now}')
        print(f'PROCESO DE UNIDADES FINALIZADA')

        # ACTUALIZAR CARGOS
        archivo_cargos = 'cargos_contratos.xlsx'
        path_file = os.path.join(SITE_ROOT, 'media', archivo_cargos)
        name_hoja = pd.ExcelFile(path_file).sheet_names[0]
        df_c = pd.read_excel(path_file, sheet_name=name_hoja)
        cont_c = 0
        for index, row in df_c.iterrows():
            cargo_old = unidecode(str(row['cargo_old']).strip().replace('  ', ' ').replace('  ', ' '))
            cargo_now = unidecode(str(row['cargo_now']).strip().replace('  ', ' ').replace('  ', ' '))
            ePersonaAcciones = PersonaContratos.objects.filter(status=True, denominacionpuesto__isnull=True, cargo__unaccent__iexact=cargo_old)
            for contrato in ePersonaAcciones:
                denominacionpuesto = DenominacionPuesto.objects.filter(status=True, descripcion__unaccent__iexact=cargo_now).first()
                if not denominacionpuesto:
                    denominacionpuesto = DenominacionPuesto(descripcion=cargo_now)
                    denominacionpuesto.save()
                    print(f'Se creo nuevo cargo {denominacionpuesto.id} | {cargo_now}')
                    # print(f'Se creo nuevo cargo, {cargo_now}')
                # contrato.cargo = cargo_now
                contrato.denominacionpuesto = denominacionpuesto
                contrato.save(update_fields=['denominacionpuesto', 'cargo'])
                cont_c += 1
                print(f'{contrato.id}) Se edito cargo de {contrato}-{contrato.persona} | Anterior={cargo_old} <----> Actual={cargo_now}')
        print(f'PROCESO DE CARGOS FINALIZADA')
        print(f'Finalizo procesos, Total unidades actualizadas: {cont_u} | Total cargos actualizados:{cont_c}')
    except Exception as ex:
        transaction.set_rollback(True)
        print(f'{ex}')

# actualizar_accionpersonal()

def actualizar_activopago_banco():
    try:
        asignados = BecaAsignacion.objects.filter(status=True, solicitud__periodo_id=317,
                                                  solicitud__becatipo_id__isnull=False)
        for a in asignados:
            person = a.solicitud.inscripcion.persona
            cuentas = person.cuentabancariapersona_set.filter(status=True)
            if len(cuentas) == 1:
                cuenta = cuentas.first()
                if cuenta.activapago == False:
                    cuenta.activapago = True
                    cuenta.save()
                print(f'se actualizo activopago true {cuenta}')
        print(f'PROCESO DE CARGOS FINALIZADA')
    except Exception as ex:
        transaction.set_rollback(True)
        print(f'{ex}')

# actualizar_activopago_banco()


def actualizar_datos_actualizadosbecas():
    try:
        asignados = BecaAsignacion.objects.filter(status=True, solicitud__periodo_id=317, infoactualizada=False,
                                                  solicitud__inscripcion__persona__personadocumentopersonal__isnull=False,
                                                  solicitud__inscripcion__persona__cuentabancariapersona__isnull=False,
                                                  solicitud__inscripcion__persona__cuentabancariapersona__activapago=True,
                                                  solicitud__becatipo_id__isnull=False).order_by('id').distinct()
        for a in asignados:
            a.infoactualizada= True
            a.save()
            print(f'se actualizo infoactualizada true {a}')
        print(f'PROCESO DE actualizacion de datos finalizada')
    except Exception as ex:
        transaction.set_rollback(True)
        print(f'{ex}')

# actualizar_datos_actualizadosbecas()


def actualizar_respuestas_pazsalvo():
    try:
        print(f'Inicia el proceso')
        respuestas = DetallePazSalvo.objects.filter(status=True, pazsalvo__formato_id=3, pazsalvo__estado=1, pazsalvo__status=True)

        for r in respuestas:
            pregunta = r.pregunta
            pregunta_new = DetalleDireccionFormatoPS.objects.filter((Q(formato_id=3) | Q(direccionformato__formato_id=3)), descripcion=pregunta.descripcion).first()
            print(pregunta_new,'-----------',r)
            if pregunta_new:
                r.pregunta = pregunta_new
                r.save(update_fields=['pregunta'])
                print(f'se actualizo respuesta true {r}')
        print(f'PROCESO DE actualizacion de datos finalizada')
    except Exception as ex:
        transaction.set_rollback(True)
        print(f'{ex}')


#actualizar_respuestas_pazsalvo()

def calcular():
    solicitudes = SolicitudPagoBeca.objects.filter(status=True, periodo_id=317)
    for s in solicitudes:
        totalsidetalle = SolicitudPagoBecaDetalle.objects.filter(solicitudpago=s, status=True)
        s.cantidadbenef = len(totalsidetalle)
        s.montopago = totalsidetalle.aggregate(valor=Sum('monto'))['valor']
        s.save()


def crear_registros_solicitudpagobeca():
    from sagest.models import Rubro
    list_ruc = ['0904451697001', '0925989758001', '0916365547001', '0920735800001', '0914192000001', '0915814917001']
    for ruc in list_ruc:
        rubros = Rubro.objects.filter(persona__ruc=ruc, status=True, cancelado=False)
        for rubro in rubros:
            if not rubro.tiene_pagos():
                rubro.iva_id = 4
                rubro.save()
                print(f"Rubro {rubro.id} actualizado")

def cambiar_estado_pazsalvo():
    from sagest.models import PazSalvo, HistorialCertificadoFirmaPS
    pazsalvos = PazSalvo.objects.filter(status=True, estado=2, formato_id=1)
    for p in pazsalvos:
        total_firmas = len(HistorialCertificadoFirmaPS.objects.filter(status=True, certificadosalida__pazsalvo=p, estado=2).values_list('persona_id', flat=True).order_by('persona_id').distinct())
        if total_firmas > 11:
            p.responsables().filter(firmado=False).update(firmado=True)
            p.estado = 3
            p.save()
            print('actualizado')


def migrar_ususariosconsulta():
    from sagest.models import UsuarioConsultaEvidencia, UsuarioEvidencia
    usuarios = UsuarioConsultaEvidencia.objects.filter(status=True)
    for usuario in usuarios:
        user = UsuarioEvidencia.objects.filter(userpermiso=usuario.userpermiso,
                                               unidadorganica=usuario.unidadorganica,
                                               tipousuario=usuario.tipousuario,
                                               tipopermiso=3).first()
        if not user:
           user = UsuarioEvidencia(userpermiso=usuario.userpermiso,
                                   unidadorganica=usuario.unidadorganica,
                                   tipousuario=usuario.tipousuario,
                                   tipopermiso=3)
           user.save()
           print(f'Usuario {user} creado')

# migrar_ususariosconsulta()

def activar_cambioclave():
    from sga.models import ProfesorMateria, Persona
    ids_personas = ProfesorMateria.objects.filter(status=True, materia__nivel__periodo_id=317, materia__status=True).\
                    values_list('profesor__persona_id',flat=True).order_by('profesor__persona_id').distinct()
    for persona in Persona.objects.filter(id__in=ids_personas):
        persona.cambiar_clave()
        print(persona)

# activar_cambioclave()


# Ejecución de una sola vez no ejecutar más de una vez -- ya fue ejecutado
def migrar_capacitacion_titulo():
    try:
        from postulate.models import PersonaCapacitacionesPartida, PersonaFormacionAcademicoPartida, PersonaAplicarPartida
        from sga.models import Persona, Capacitacion
        persona = Persona.objects.get(id=328233)
        capacitacion = Capacitacion.objects.get(id=78177)
        aplicar = PersonaAplicarPartida.objects.filter(persona=persona, status=True).first()
        instancia = PersonaCapacitacionesPartida(personapartida=aplicar, institucion=capacitacion.institucion, tipo=capacitacion.tipo,
                                                 nombre=capacitacion.nombre, descripcion=capacitacion.descripcion, tipocurso=capacitacion.tipocurso,
                                                 tipocapacitacion=capacitacion.tipocapacitacion, tipocertificacion=capacitacion.tipocertificacion, tipoparticipacion=capacitacion.tipoparticipacion,
                                                 anio=capacitacion.anio, contextocapacitacion=capacitacion.contextocapacitacion, detallecontextocapacitacion=capacitacion.detallecontextocapacitacion,
                                                 auspiciante=capacitacion.auspiciante, areaconocimiento=capacitacion.areaconocimiento, subareaconocimiento=capacitacion.subareaconocimiento,
                                                 subareaespecificaconocimiento=capacitacion.subareaespecificaconocimiento,
                                                 pais=capacitacion.pais, provincia=capacitacion.provincia, canton=capacitacion.canton, parroquia=capacitacion.parroquia, fechainicio=capacitacion.fechainicio, fechafin=capacitacion.fechafin,
                                                 horas=capacitacion.horas,
                                                 expositor=capacitacion.expositor, modalidad=capacitacion.modalidad, otramodalidad=capacitacion.otramodalidad)
        instancia.archivo = capacitacion.archivo if capacitacion.archivo else None
        instancia.save()

        mi_f = persona.mi_formacionacademica().last()
        formacion = PersonaFormacionAcademicoPartida.objects.get(id=18143)
        formacion.personapartida = aplicar
        formacion.titulo = mi_f.titulo
        formacion.registro = mi_f.registro
        formacion.pais = mi_f.pais
        formacion.provincia = mi_f.provincia
        formacion.canton = mi_f.canton
        formacion.parroquia = mi_f.parroquia
        formacion.educacionsuperior = mi_f.educacionsuperior
        formacion.institucion = mi_f.institucion
        formacion.cursando = mi_f.cursando
        formacion.archivo = mi_f.archivo if mi_f.archivo else None
        formacion.save()
        if CamposTitulosPostulacion.objects.filter(status=True, titulo=mi_f.titulo).exists():
            campotitulo = CamposTitulosPostulacion.objects.filter(status=True, titulo=mi_f.titulo).first()
            formacion.camposamplio.clear()
            formacion.campoespecifico.clear()
            formacion.campodetallado.clear()
            for ct in campotitulo.campoamplio.all():
                formacion.campoamplio.add(ct)
            for ct in campotitulo.campoespecifico.all():
                formacion.campoespecifico.add(ct)
            for ct in campotitulo.campodetallado.all():
                formacion.campodetallado.add(ct)
            formacion.save()
    except Exception as e:
        transaction.set_rollback(True)
def migrar_experiencia():
    from postulate.models import PersonaExperienciaPartida, PersonaFormacionAcademicoPartida, PersonaAplicarPartida
    from sga.models import Persona, Capacitacion
    persona = Persona.objects.get(id=69710)
    aplicar = PersonaAplicarPartida.objects.get(id=8258)
    experiencia = persona.mis_experienciaslaborales().last()
    instancia = PersonaExperienciaPartida.objects.get(id=25921)
    instancia.personapartida = aplicar
    instancia.institucion = experiencia.institucion
    instancia.actividadlaboral = experiencia.actividadlaboral
    instancia.cargo = experiencia.cargo
    instancia.fechainicio = experiencia.fechainicio
    instancia.fechafin = experiencia.fechafin
    instancia.archivo = experiencia.archivo if experiencia.archivo else None
    instancia.save()
def migrar_experiencia_sojos():
    from postulate.models import PersonaExperienciaPartida, PersonaFormacionAcademicoPartida, PersonaAplicarPartida
    from sga.models import Persona, Capacitacion
    persona = Persona.objects.get(id=38455)
    aplicar = PersonaAplicarPartida.objects.get(id=8251)
    experiencia = persona.mis_experienciaslaborales().last()
    instancia = PersonaExperienciaPartida.objects.get(id=25899)
    instancia.personapartida = aplicar
    instancia.institucion = experiencia.institucion
    instancia.actividadlaboral = experiencia.actividadlaboral
    instancia.cargo = experiencia.cargo
    instancia.fechainicio = experiencia.fechainicio
    instancia.fechafin = experiencia.fechafin
    instancia.archivo = experiencia.archivo if experiencia.archivo else None
    instancia.save()
def migrar_experiencia_add_sojos():
    from postulate.models import PersonaExperienciaPartida, PersonaAplicarPartida
    from sga.models import Persona
    from sagest.models import ExperienciaLaboral
    persona = Persona.objects.get(id=38455)
    aplicar = PersonaAplicarPartida.objects.get(id=8251)
    experiencia = ExperienciaLaboral.objects.get(id=34717)
    instancia = PersonaExperienciaPartida(personapartida=aplicar,
                                          institucion=experiencia.institucion,
                                          actividadlaboral=experiencia.actividadlaboral,
                                          cargo=experiencia.cargo,
                                          fechainicio=experiencia.fechainicio,
                                          fechafin=experiencia.fechafin,
                                          archivo=experiencia.archivo)
    instancia.save()
def capacitacion_doc_jara():
    from postulate.models import PersonaCapacitacionesPartida
    from sga.models import Persona, Capacitacion
    persona = Persona.objects.get(id=69710)
    capacitacion = persona.mis_capacitaciones().last()
    instancia = PersonaCapacitacionesPartida.objects.get(id=56431)
    instancia.archivo = capacitacion.archivo if capacitacion.archivo else None
    instancia.save()

def activar_cambioclave():
    from sga.models import ProfesorMateria, Persona
    from sagest.models import DistributivoPersona
    ids_personas = ProfesorMateria.objects.filter(status=True,
                                                  materia__nivel__periodo_id=397,
                                                  materia__status=True).values_list('profesor__persona_id', flat=True).order_by('profesor__persona_id').distinct()

    # Docentes, administrativos y trabajadores
    distributivos = DistributivoPersona.objects.filter((Q(regimenlaboral_id__in=[1, 4]) | Q(persona_id__in=ids_personas)), status=True)
    for d in distributivos:
        d.persona.cambiar_clave()
        print(d.persona)

def cambiopass_required():
    from sagest.models import DistributivoPersona
    # Todos los de regimen laboral losep, loes y codigo de trabajo
    distributivos = DistributivoPersona.objects.filter(regimenlaboral_id__in=[1, 2, 4], status=True)
    for d in distributivos:
        d.persona.cambiar_clave()
        print(d.persona)
# cambiopass_required()
def agregar_titulos_partida():
    from postulate.models import Partida
    from sga.models import Titulo
    partida = Partida.objects.get(id=863)
    titulos = Titulo.objects.filter(status=True, id__in=[11200, 4118])
    for tit in titulos:
        partida.titulos.add(tit)


def marcar_individual():
    from sagest.models import HistorialJornadaTrabajador, LogMarcada, LogDia
    from sga.models import Persona
    cedula = '0302244926'
    if cedula:
        servidor = Persona.objects.filter(Q(cedula=cedula)|Q(pasaporte=cedula),status=True)
        if servidor:
            servidor = servidor[0]
            hoy = datetime.now()
            fecha = date(2024, 1, 24)
            secuencia = 4
            jornada = HistorialJornadaTrabajador.objects.filter(status=True, persona=servidor,
                                                                fechainicio__lte=hoy,
                                                                fechafin__isnull=True).first()
            if jornada:
                detalles = jornada.jornada.detalle_jornada()
                detalles = detalles.filter(dia=2)
                logdia = LogDia.objects.filter(status=True, fecha=fecha, persona=servidor).first()
                if not logdia:
                    logdia = LogDia(fecha=fecha, persona=servidor, jornada=jornada.jornada)
                    logdia.save()
                marcadas_old = LogMarcada.objects.filter(logdia=logdia)
                marcadas_old.delete()
                for deta in detalles:
                    fin = datetime(fecha.year, fecha.month, fecha.day, deta.horafin.hour, deta.horafin.minute,
                                   deta.horafin.second)
                    inicio = datetime(fecha.year, fecha.month, fecha.day, deta.horainicio.hour,
                                      deta.horainicio.minute, deta.horainicio.second)
                    secuencia += 1
                    marcada_inicio = LogMarcada(status=True, logdia=logdia, time=inicio, secuencia=secuencia, similitud=98.880)
                    marcada_inicio.save()
                    secuencia += 1
                    marcada_fin = LogMarcada(status=True, logdia=logdia, time=fin, secuencia=secuencia, similitud=99.780)
                    marcada_fin.save()
                total = logdia.logmarcada_set.all().count()
                logdia.cantidadmarcadas = total
                logdia.save()