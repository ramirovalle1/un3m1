# coding=utf-8
from __future__ import division

import sys
import uuid
import random
import string
from hashlib import md5

from django.core.exceptions import ObjectDoesNotExist
from django.http import JsonResponse, HttpResponseRedirect
from django_user_agents.utils import get_user_agent
from bd.models import UserToken, UserAccessSecurity, UserAccessSecurityType, UserAccessSecurityDevice, \
    UserAccessSecurityCode,InventarioOpcionSistema
from secrets import token_hex

from core.cache import get_cache_ePerfilUsuario
from settings import EMAIL_DOMAIN, DEBUG
from sga.funciones import log, variable_valor, logquery
from sga.models import miinstitucion, CUENTAS_CORREOS, Modulo, Periodo, PerfilInscripcion,Persona
from sga.tasks import send_html_mail
from sga.templatetags.sga_extras import encrypt
from datetime import datetime, timedelta
from django.contrib.contenttypes.models import ContentType
from django.db import transaction, connections
unicode = str


def recoveryPassword(request, persona, app):
    from sga.datosiniciales import generar_cambio_clave
    with transaction.atomic():
        try:
            rescedula = variable_valor('RESET_CONTRASENA_CEDULA')
            clave_tmp = token_hex(16)
            fecha = datetime.now().date()
            hora = datetime.now().time()
            fecha_hora = fecha.year.__str__() + fecha.month.__str__() + fecha.day.__str__() + hora.hour.__str__() + hora.minute.__str__() + hora.second.__str__()
            user = persona.usuario
            if not user:
                raise NameError(u"Usuario no encontrado")
            usuario = persona.usuario
            if rescedula:
                usuario.set_password(persona.identificacion())
                mensajes = "Su nueva clave es su número de cédula"
            else:
                usuario.set_password(clave_tmp)
                mensajes = "Revise su bandeja de correo para completar la solicitud"
            usuario.save()
            token = md5(str(encrypt(usuario.id)+fecha_hora).encode("utf-8")).hexdigest()
            cambioclave = persona.cambiar_clave()
            cambioclave.solicitada = True
            cambioclave.clavecambio = generar_cambio_clave()
            cambioclave.save(request)
            log(u'Genero nuevo cambio de contraseña: %s [%s]' % (persona, persona.id), request, "add", user=user)
            if UserToken.objects.filter(user=user, isActive=True, action_type=1, date_expires__gte=datetime.now()).exists():
                raise NameError(u"Mantiene una recuperación de contraseña activa.")
            else:
                UserToken.objects.filter(user=user, isActive=True, action_type=1).update(isActive=False)
                eUserToken = UserToken(user=usuario,
                                       token=token,
                                       action_type=1,
                                       date_expires=datetime.now() + timedelta(days=1),
                                       app=app,
                                       isActive=True)
                eUserToken.save(request)
                app_label = ''
                sistema = ''
                if app == 1:
                    app_label = 'sga'
                    sistema = u'Sistema de Gestión Académica'
                elif app == 2:
                    app_label = 'sagest'
                    sistema = u'Sistema de Gestión Administrativa'
                elif app == 3:
                    app_label = 'admisionposgrado'
                    sistema = u'Sistema de Gestión Académica'
                elif app == 4:
                    app_label = 'sgaestudiante'
                    sistema = u'Sistema de Gestión Académica Estudiantil'
                else:
                    app_label = 'sga'
                    sistema = u'Sistema de Gestión Académica'
                if not DEBUG:
                    send_html_mail("Solicitud de cambio de contraseña", "emails/cambioclave_new.html",
                                   {'sistema': sistema,
                                    'fecha': datetime.now().date(),
                                    'persona': persona,
                                    'token': token,
                                    'app_label': app_label,
                                    'app': app,
                                    't': miinstitucion(),
                                    'dominio': EMAIL_DOMAIN
                                    },
                                   persona.lista_emails(), [],
                                   cuenta=CUENTAS_CORREOS[7][1])
            return JsonResponse({"result": "ok", "mensaje": f'{mensajes}'})

        except Exception as ex:
            transaction.set_rollback(True)
            return JsonResponse({"result": "bad", "mensaje": u"Error, %s" % ex.__str__()})


def generate_code(max=6):
    code_str = string.ascii_letters + string.digits
    code = ''.join(random.sample(code_str, max))
    return code


def get_client_ip(request):
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip


def obtener_optionSystem(url):
    try:
        return InventarioOpcionSistema.objects.values_list('id', 'nombre', 'descripcion', 'modulo', 'modulo_id').filter(status=True, url=url).last()
    except Exception as ex:
        return InventarioOpcionSistema.objects.none()


def enviar_notificacion_log_store_uxplora(datos):
    import json
    from channels.layers import get_channel_layer
    from asgiref.sync import async_to_sync

    channel_layer = get_channel_layer()
    async_to_sync(channel_layer.group_send)(
        'room_interface_uxplora',
        {
            'type': 'dio_clic',
            'datos': json.dumps(datos),
            'message': 'no_se'
        }
    )


def calcular_edad(fecha_nacimiento):
    from datetime import date
    fecha_actual = date.today()
    edad = fecha_actual.year - fecha_nacimiento.year
    if fecha_actual.month < fecha_nacimiento.month or (fecha_actual.month == fecha_nacimiento.month and fecha_actual.day < fecha_nacimiento.day):
        edad -= 1
    return edad


def log_store(instancia, action, log, crud, request, eOrigin, periodo, materia=None, relateduser=None, other=None, inscripcion=None,perfil=None,url_clic=None,modulo=None):
    eIp = get_client_ip(request)
    eAction = action.id if action else "Null"
    componente = action.nombre if action else ''
    evento = action.descripcion if action else ''
    eObjecttable = ContentType.objects.get_for_model(instancia) if instancia else ''
    eObjectid = instancia.id if instancia else "Null"
    eUser = request.user.id if request.user else "Null"
    eMateria = materia if materia else "Null"
    eRelateduser = relateduser.id if relateduser else "Null"
    eOther = other if other else ''
    eRealuser = request.user.id if request.user else "Null"
    eNse = "Null"
    eEdulevel = "Null"
    eMatricula = "Null"
    eInscripcion = "Null"
    matricula = "Null"
    eCarrera = "Null"
    eFacultad = "Null"
    if inscripcion:
        eInscripcion = inscripcion.id if inscripcion else "Null"
        eCarrera = inscripcion.carrera.nombre if inscripcion.carrera else "Null"
        if inscripcion.carrera.mi_coordinacion():
            eFacultad = inscripcion.carrera.mi_coordinacion() if inscripcion.carrera else "Null"
        if inscripcion.mi_matricula_periodo(periodo) and periodo:
            matricula = inscripcion.mi_matricula_periodo(periodo)
            eMatricula = matricula.id
            eEdulevel = matricula.nivel.id
            if matricula.matriculagruposocioeconomico_set.values('id').filter(status=True).exists():
                eNse = matricula.matriculagruposocioeconomico_set.filter(status=True)[0].gruposocioeconomico.nombre
    ePersona=Persona.objects.get(usuario=eUser)
    eNombres=ePersona.nombre_completo_inverso() if ePersona.nombre_completo_inverso() else "Null"
    eIdentificacion=ePersona.identificacion() if ePersona.identificacion() else "Null"
    eFechanacimiento=ePersona.nacimiento if ePersona.nacimiento else "Null"
    eEdad=calcular_edad(ePersona.nacimiento) if ePersona.nacimiento else "Null"
    eNombreperiodo=periodo.nombre if periodo else "Null"
    ePerfilInscripcion = PerfilInscripcion.objects.filter(persona=ePersona)
    eEtnia="Null"
    eTipodiscapacidad="Null"
    if ePerfilInscripcion.values('id').exists():
        eTipodiscapacidad =ePerfilInscripcion[0].tipodiscapacidad.nombre if ePerfilInscripcion[0].tipodiscapacidad else "Null"
        eEtnia = ePerfilInscripcion[0].raza.nombre if ePerfilInscripcion[0].raza else "Null"
    ePais=ePersona.pais.nombre if ePersona.pais else "Null"
    eProvincia=ePersona.provincia.nombre if ePersona.provincia else "Null"
    eCanton=ePersona.canton.nombre if ePersona.canton else "Null"
    elgtbi=ePersona.lgtbi
    eSexo=ePersona.sexo.nombre if ePersona.sexo else  "Null"
    eContextid = periodo.id if periodo else "Null"
    nombremodulo = Modulo.objects.get(id=modulo).nombre if modulo else "Null"
    nombreopcion = action.nombre if action else "Null"
    with transaction.atomic():
        try:
            conexion = connections['uxplora']
            cursor = conexion.cursor()
            sql = """ INSERT INTO logstore_logstoresystem (status, eventname, component, action, module, target, objecttable, 
                                                            objectid, crud, edulevel, contextid, userid, course, relateduser, other, 
                                                            origin, ip, realuser, fecha_creacion, inscripcion, matricula,perfil,url,
                                                            nombres,identificacion,fechanacimiento,edad,nombreperiodo,facultad,carrera,
                                                            tipodiscapacidad,nse,etnia,pais,provincia,canton,lgtbi,sexo,nombremodulo,nombreopcion) 
                                                              VALUES (TRUE, '%s', '%s', %s, '%s','%s','%s',
                                                               %s, '%s', %s, %s, %s, %s, %s, '%s', 
                                                              '%s','%s', %s, NOW(), %s,  %s, %s, '%s',
                                                              '%s', '%s', '%s', %s, '%s', '%s', '%s',
                                                              '%s', '%s', '%s', '%s', '%s', '%s', '%s','%s','%s','%s'); """ \
                                                          % (evento, componente, eAction, modulo, log, eObjecttable,
                                                             eObjectid, crud, eEdulevel, eContextid, eUser, eMateria, eRelateduser, eOther,
                                                             eOrigin, eIp, eRealuser, eInscripcion, eMatricula, perfil,url_clic.replace("<ASGIRequest:","").replace("'","").replace(">",""),
                                                             eNombres,eIdentificacion,eFechanacimiento,eEdad,eNombreperiodo,eFacultad,eCarrera,
                                                             eTipodiscapacidad,eNse,eEtnia,ePais,eProvincia,eCanton,elgtbi,eSexo, nombremodulo, nombreopcion  )
            cursor.execute(sql)
        except Exception as ex:
            textoerror = '{} Linea:{}'.format(ex, sys.exc_info()[-1].tb_lineno)
            transaction.set_rollback(True, using='uxplora')
        finally:
            cursor.close()


def action_registre_log_clic_commonviews(request):
    try:
        if request.path[1:]:
            perfilprincipal = request.session['perfilprincipal']
            if perfilprincipal.es_profesor() or perfilprincipal.es_estudiante():
                data = {}
                urlPath = None
                data['usuario_id'] = request.user.id
                data['usuario'] = request.user.username
                data['fecha'] = u"%s"%datetime.now().date()
                data['hora'] = (datetime.now().time()).strftime("%H:%M:%S")
                data['perfil'] =perfil= 1 if perfilprincipal.es_estudiante() else 2
                eAction = request.GET.get('action', None)
                eInventarioOpcionSistema=None
                urlaction=None
                urlauxiliar=None
                if eAction is None:
                    eAction = request.POST.get('action', None)
                if eAction:
                    data['action'] = eAction
                    if '/' in request.path[1:]:
                        sp = request.path[1:].split('/')
                        urlPath = sp[0]
                    else:
                        urlPath = request.path[1:]
                    urlaction = urlPath + '?action=' + eAction
                else:
                    if '/' in request.path[1:]:
                        sp = request.path[1:].split('/')
                        urlPath = sp[0]
                        urlaction = urlPath
                    elif '?' in str(request):
                        urlPath = request.path[1:]
                        urlauxiliar = str(request).replace("<ASGIRequest:","").replace("'","").replace(">","")
                    else:
                        urlPath = request.path[1:]
                        urlaction = urlPath
                data['url_path'] = urlaction if eAction else urlauxiliar if urlauxiliar else urlPath
                try:
                    eModulo = Modulo.objects.get(url=urlPath)
                    data['modulo_id'] =eModuloid= eModulo.id
                except ObjectDoesNotExist:
                    eModulo = Modulo.objects.none()
                instancia, materia, inscripcion, relateduser = None, None, None, None
                try:
                    eInventarioOpcionSistema = InventarioOpcionSistema.objects.get(status=True, url=f"/{urlaction}")
                except ObjectDoesNotExist:
                    eInventarioOpcionSistema = InventarioOpcionSistema(url=f"/{urlaction}",modulo=eModulo,proceso_id=81,nombre='SN',descripcion='SN')
                    eInventarioOpcionSistema.save()
                log = u'Usuario %s (%s) con perfil %s %s módulo %s.' % (
                    perfilprincipal.persona, request.user, perfilprincipal,
                    'ejecutó la acción ' + eAction + ' en el' if eAction else 'accedió al', eModulo.nombre)
                crud = request.method
                eOrigin = request.session['nombresistema'] if request.session['nombresistema'] else ''
                periodo = request.session['periodo']
                if perfilprincipal.es_estudiante():
                    inscripcion = perfilprincipal.inscripcion
                # other = str(request.user_agent)  # Nombre del dispositivo en el que accede
                other = 'PC / Windows 10 / Chrome 117.0.0'  # Datos del dispositivo en el que accede
                log_store(instancia,
                          eInventarioOpcionSistema,
                          log,
                          crud,
                          request,
                          eOrigin,
                          periodo,
                          materia,
                          relateduser,
                          other,
                          inscripcion,
                          perfil,
                          urlaction if eAction else urlauxiliar if urlauxiliar else urlPath,
                          eModuloid)
                enviar_notificacion_log_store_uxplora(data)

    except Exception as ex:
        textoerror = '{} Linea:{}'.format(ex, sys.exc_info()[-1].tb_lineno)
        pass


def action_registre_log_clic_api_sge(request, view):
    try:
        payload = request.auth.payload
        ePerfilUsuario = get_cache_ePerfilUsuario(int(encrypt(payload['perfilprincipal']['id'])))
        data = {}
        urlaction = None
        data['usuario_id'] = request.user.id
        data['usuario'] = request.user.username
        data['fecha'] = u"%s" % datetime.now().date()
        data['hora'] = (datetime.now().time()).strftime("%H:%M:%S")
        data['perfil'] = perfil=1
        eAction = request.query_params.get('action', None)  # GET
        urlPath=None
        if eAction is None:
            eAction = request.data.get('action', None)  # POST
        if eAction:
            data['action'] = eAction
        try:
            eModulo = Modulo.objects.get(api_key=view.api_key_module)
        except ObjectDoesNotExist:
            eModulo = Modulo.objects.none()
        if eModulo:
            data['modulo_id'] =modulo_id= eModulo.id
            urlaction = eModulo.url
            if eAction:
                urlaction = eModulo.url + '?action=' + eAction
            urlPath=urlaction
        else:
            urlPath = str(request)
        data['url_path'] = urlaction if eAction else urlPath
        instancia, materia, inscripcion, relateduser = None, None, None, None
        try:
            eInventarioOpcionSistema = InventarioOpcionSistema.objects.get(status=True, url=f"/{urlaction}")
        except ObjectDoesNotExist:
            eInventarioOpcionSistema = InventarioOpcionSistema(url=f"/{urlaction}", modulo=eModulo,proceso_id=81, nombre='SN',
                                                               descripcion='SN')
            eInventarioOpcionSistema.save()
        log = u'Usuario %s (%s) con perfil %s %s módulo %s.' % (
        ePerfilUsuario.persona, request.user, ePerfilUsuario,
        'ejecutó la acción ' + eAction + ' en el' if eAction else 'accedió al', eModulo.nombre if eModulo else '')
        crud = request.method
        eOrigin = payload['templatebasesetting']['name_system']
        periodo = Periodo.objects.get(pk=int(encrypt(payload['periodo']['id'])))
        # if perfilprincipal.es_estudiante():
        inscripcion = ePerfilUsuario.inscripcion
        # other = str(request.user_agent)  # Nombre del dispositivo en el que accede
        other = 'PC / Windows 10 / Chrome 117.0.0' # Nombre del dispositivo en el que accede
        log_store(instancia,
                  eInventarioOpcionSistema,
                  log,
                  crud,
                  request,
                  eOrigin,
                  periodo,
                  materia,
                  relateduser,
                  other,
                  inscripcion,perfil,urlaction if urlaction else urlPath,
                  modulo_id)
        enviar_notificacion_log_store_uxplora(data)
    except Exception as ex:
        pass


def export_query_to_excel(request, query, cursorconection, nombre_archivo, params, full_name, baseafectada, sheet_name='Reporte'):
    import pandas as pd
    from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    rows_effected = 0
    try:
        with cursorconection as cursor:
            cursor.execute(query, params)
            columns = [col[0] for col in cursor.description]
            data = cursor.fetchall()
            rows_effected = cursor.rowcount

        df = pd.DataFrame(data, columns=columns)

        with pd.ExcelWriter(full_name, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name=sheet_name)

            workbook = writer.book
            worksheet = writer.sheets[sheet_name]

            # Define styles
            header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
            border_style = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            header_font = Font(bold=True)
            alignment = Alignment(horizontal='center', vertical='center')

            # Apply styles to header
            for col_num, column_title in enumerate(df.columns, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = alignment
                cell.border = border_style
                # Adjust column width
                worksheet.column_dimensions[get_column_letter(col_num)].width = 20

            # Apply styles to all cells
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=len(df.columns)):
                for cell in row:
                    cell.border = border_style
        url = f'/media/reportes/gestion/{nombre_archivo}'
        logquery(baseafectada, query, rows_effected, request, url=url)
        res_json = {'error': False, 'url': url}
        return res_json
    except Exception as e:
        res_json = {'error': True, 'message': f'{e}'}
        return res_json


