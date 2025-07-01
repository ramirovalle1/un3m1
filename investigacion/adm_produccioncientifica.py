# -*- coding: UTF-8 -*-
import io
import json
import os
import zipfile
from math import ceil

import xlsxwriter
import openpyxl
from django.db.models.functions import ExtractYear

import PyPDF2
from datetime import time
from decimal import Decimal
from datetime import datetime

import requests
from dateutil.relativedelta import relativedelta
from django.contrib.auth.decorators import login_required
from django.core.files.base import ContentFile
from django.db import transaction
from django.db.models import Q
from django.forms import model_to_dict
from django.http import HttpResponseRedirect, JsonResponse, HttpResponse
from django.shortcuts import render
import time as pausaparaemail
from django.core.files import File as DjangoFile
from fitz import fitz
from xlwt import easyxf, XFStyle, Workbook
import random

from core.firmar_documentos_ec import JavaFirmaEc
from decorators import secure_module
from investigacion.forms import GrupoInvestigacionForm, EvidenciasCodigoForm
from investigacion.funciones import coordinador_investigacion, tecnico_revisor_grupoinvestigacion, notificar_grupo_investigacion, vicerrector_investigacion_posgrado, experto_investigacion, secuencia_informe_grupoinvestigacion, responsable_coordinacion, FORMATOS_CELDAS_EXCEL, \
    notificar_revision_solicitud_produccion_cientifica, codigo_publicacion_valido
from investigacion.models import GrupoInvestigacion, FUNCION_INTEGRANTE_GRUPO_INVESTIGACION, GrupoInvestigacionIntegrante, GrupoInvestigacionRecorrido, GrupoInvestigacionIntegranteRequisito, GrupoInvestigacionInforme, GrupoInvestigacionInformeAnexo, GrupoInvestigacionResolucion
from sagest.commonviews import obtener_estado_solicitud, obtener_estados_solicitud
from sagest.models import SolicitudPublicacion
from settings import SITE_STORAGE
from sga.commonviews import adduserdata
from sga.funciones import MiPaginador, log, variable_valor, \
    convertir_fecha, remover_caracteres_tildes_unicode, cuenta_email_disponible, generar_nombre, email_valido, \
    validar_archivo, null_to_decimal, cuenta_email_disponible_para_envio, fechaletra_corta, remover_atributo_style_html, remover_caracteres, elimina_tildes, remover_caracteres_especiales_unicode
from sga.funcionesxhtml2pdf import convert_html_to_pdf
from sga.models import CUENTAS_CORREOS, Persona, MESES_CHOICES, ArticuloInvestigacion, ParticipantesArticulos, ArticulosBaseIndexada, DetalleEvidencias
from django.template import Context
from django.template.loader import get_template

import time as ET

from sga.tasks import send_html_mail, conectar_cuenta
from sga.templatetags.sga_extras import encrypt

from django.core.cache import cache
from django.contrib import messages

@login_required(redirect_field_name='ret', login_url='/loginsga')
@transaction.atomic()
@secure_module
def view(request):
    data = {}
    adduserdata(request, data)
    data['persona'] = persona = request.session['persona']
    perfilprincipal = request.session['perfilprincipal']
    es_administrativo = perfilprincipal.es_administrativo()

    if not es_administrativo:
        return HttpResponseRedirect("/?info=El Módulo está disponible para administrativos.")

    if request.method == 'POST':
        action = request.POST['action']

        if action == 'evidenciasarticulo':
            try:
                if not request.POST['anioevidencia']:
                    return JsonResponse({"result": "bad", "mensaje": u"Seleccione el año de publicación", "typewarning": "typewarning"})

                import time as ET
                # Aquí se almacenan las fichas catalográficas
                directorio = SITE_STORAGE + '/media/articulos'
                try:
                    os.stat(directorio)
                except:
                    os.mkdir(directorio)

                caracteres_a_quitar = "!\"\#$%&()=?¡'¿{}[],.-+*/|°^~:;"
                anio = request.POST['anioevidencia']
                output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media', 'zipav'))

                response = HttpResponse(content_type='application/zip')
                response['Content-Disposition'] = 'attachment; filename=articulosevidencias_' + random.randint(1, 10000).__str__() + '.zip'

                nombre_archivo = "evidenciasarticulos" + anio + ".zip"
                filename = os.path.join(output_folder, nombre_archivo)
                fantasy_zip = zipfile.ZipFile(filename, 'w')

                articulos_publicados = ArticuloInvestigacion.objects.filter(status=True, fechapublicacion__year=anio).order_by('id')
                total1 = articulos_publicados.count()

                articulos_no_publicados = ArticuloInvestigacion.objects.filter(status=True, fechapublicacion__isnull=True, fechaaprobacion__year=anio).order_by('id')
                total2 = articulos_no_publicados.count()

                total = total1 + total2
                rp = 0
                # articulos_publicados = ArticuloInvestigacion.objects.filter(status=True, fechapublicacion__year=anio).order_by('id')[:10]

                for articulo in articulos_publicados:
                    rp += 1
                    print("Procesando", rp, " de ", total)
                    titulo = articulo.nombre

                    palabras = titulo.split(" ")
                    titulo = "_".join(palabras[0:5])
                    titulo = remover_caracteres(titulo, caracteres_a_quitar)

                    titulo = elimina_tildes(remover_caracteres_tildes_unicode(remover_caracteres_especiales_unicode(titulo)))
                    carpetaarticulo = "PUBLICADOS/ART_" + articulo.revista.codigoissn.strip().replace('-', '_') + "_" + str(articulo.id)+ "_" + titulo

                    articulo_id = articulo.id
                    nombre = "ARTICULO_" + str(articulo_id).zfill(4)

                    # Agrego las evidencias a la carpeta del artículo
                    for evidencia in articulo.detalleevidencias_set.filter(status=True):
                        ext = evidencia.archivo.__str__()[evidencia.archivo.__str__().rfind("."):]

                        if evidencia.descripcion:
                            nombreevidencia = evidencia.descripcion.upper()
                            nombreevidencia = remover_caracteres(nombreevidencia, caracteres_a_quitar)
                            nombreevidencia = elimina_tildes(remover_caracteres_tildes_unicode(remover_caracteres_especiales_unicode(nombreevidencia))).replace(" ", "_")
                        else:
                            nombreevidencia = "EVIDENCIA" + random.randint(1, 10000).__str__()

                        if evidencia.evidencia.id == 1:
                            if os.path.exists(SITE_STORAGE + evidencia.archivo.url):
                                fantasy_zip.write(SITE_STORAGE + evidencia.archivo.url, carpetaarticulo + "/" + nombreevidencia + ext.lower())
                        elif evidencia.evidencia.id == 3:
                            if os.path.exists(SITE_STORAGE + evidencia.archivo.url):
                                fantasy_zip.write(SITE_STORAGE + evidencia.archivo.url, carpetaarticulo + "/" + nombreevidencia + ext.lower())

                    # Generar la ficha catalográfica del artículo
                    data['participantess'] = participantes = ParticipantesArticulos.objects.filter(articulo=articulo, status=True)
                    data['cantidad'] = participantes.count()
                    data['basesindexadas'] = ArticulosBaseIndexada.objects.filter(articulo=articulo, status=True)

                    nombrearchivoficha = 'fichacatalograficaart_' + str(articulo_id) + '_' + str(random.randint(1, 10000)) + '.pdf'
                    valida = convert_html_to_pdf(
                        'inv_articulos/fichacatalografica_articulo_pdf.html',
                        {'pagesize': 'A4', 'data': data},
                        nombrearchivoficha,
                        directorio
                    )

                    archivoficha = directorio + "/" + nombrearchivoficha
                    # Agrego el archivo de la ficha a la carpeta del artículo
                    fantasy_zip.write(archivoficha, carpetaarticulo + "/FICHA_CATALOGRAFICA.pdf")
                    # Borro el archivo de la ficha creado
                    ET.sleep(1)
                    os.remove(archivoficha)


                articulos_no_publicados = ArticuloInvestigacion.objects.filter(status=True, fechapublicacion__isnull=True, fechaaprobacion__year=anio).order_by('id')
                # articulos_no_publicados = ArticuloInvestigacion.objects.filter(status=True, fechapublicacion__isnull=True, fechaaprobacion__year=anio).order_by('id')[:10]
                for articulo in articulos_no_publicados:
                    rp += 1
                    print("Procesando", rp, " de ", total)

                    titulo = articulo.nombre

                    palabras = titulo.split(" ")
                    titulo = "_".join(palabras[0:5])
                    titulo = remover_caracteres(titulo, caracteres_a_quitar)

                    titulo = elimina_tildes(remover_caracteres_tildes_unicode(remover_caracteres_especiales_unicode(titulo)))
                    carpetaarticulo = "ACEPTADOPUBLICACION/ART_" + articulo.revista.codigoissn.strip().replace('-', '_') + "_" + str(articulo.id)+ "_" + titulo

                    articulo_id = articulo.id
                    nombre = "ARTICULO_" + str(articulo_id).zfill(4)

                    # Agrego las evidencias a la carpeta del artículo
                    for evidencia in articulo.detalleevidencias_set.filter(status=True):
                        ext = evidencia.archivo.__str__()[evidencia.archivo.__str__().rfind("."):]

                        if evidencia.descripcion:
                            nombreevidencia = evidencia.descripcion.upper()
                            nombreevidencia = remover_caracteres(nombreevidencia, caracteres_a_quitar)
                            nombreevidencia = elimina_tildes(remover_caracteres_tildes_unicode(remover_caracteres_especiales_unicode(nombreevidencia))).replace(" ", "_")
                        else:
                            nombreevidencia = "EVIDENCIA" + random.randint(1, 10000).__str__()

                        if evidencia.evidencia.id == 1:
                            if os.path.exists(SITE_STORAGE + evidencia.archivo.url):
                                fantasy_zip.write(SITE_STORAGE + evidencia.archivo.url, carpetaarticulo + "/" + nombreevidencia + ext.lower())
                        elif evidencia.evidencia.id == 3:
                            if os.path.exists(SITE_STORAGE + evidencia.archivo.url):
                                fantasy_zip.write(SITE_STORAGE + evidencia.archivo.url, carpetaarticulo + "/" + nombreevidencia + ext.lower())

                    # Generar la ficha catalográfica del artículo
                    data['participantess'] = participantes = ParticipantesArticulos.objects.filter(articulo=articulo, status=True)
                    data['cantidad'] = participantes.count()
                    data['basesindexadas'] = ArticulosBaseIndexada.objects.filter(articulo=articulo, status=True)

                    nombrearchivoficha = 'fichacatalograficaart_' + str(articulo_id) + '_' + str(random.randint(1, 10000)) + '.pdf'
                    valida = convert_html_to_pdf(
                        'inv_articulos/fichacatalografica_articulo_pdf.html',
                        {'pagesize': 'A4', 'data': data},
                        nombrearchivoficha,
                        directorio
                    )

                    archivoficha = directorio + "/" + nombrearchivoficha
                    # Agrego el archivo de la ficha a la carpeta del artículo
                    fantasy_zip.write(archivoficha, carpetaarticulo + "/FICHA_CATALOGRAFICA.pdf")
                    # Borro el archivo de la ficha creado
                    ET.sleep(1)
                    os.remove(archivoficha)


                fantasy_zip.close()

                ruta = "media/zipav/" + nombre_archivo

                return JsonResponse({'to': ruta})
            except Exception as ex:
                msg = ex.__str__()
                return JsonResponse({"result": "bad", "mensaje": u"Error al descargar evidencias. Detalle: %s" % (msg)})

        elif action == 'evidenciasarticulocodigo':
            try:
                archivo = request.FILES['archivo']
                descripcionarchivo = 'Archivo Excel de Códigos'
                resp = validar_archivo(descripcionarchivo, archivo, ['xls', 'xlsx'], '4MB')
                if resp['estado'] != "OK":
                    raise NameError(resp["mensaje"])

                import time as ET

                # Aquí se almacenan las fichas catalográficas
                directorio = SITE_STORAGE + '/media/articulos'
                try:
                    os.stat(directorio)
                except:
                    os.mkdir(directorio)

                caracteres_a_quitar = "!\"\#$%&()=?¡'¿{}[],.-+*/|°^~:;"
                output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media', 'zipav'))

                response = HttpResponse(content_type='application/zip')
                response['Content-Disposition'] = 'attachment; filename=articulosevidencias_' + random.randint(1, 10000).__str__() + '.zip'

                nombre_archivo = "evidenciasarticuloscodigo.zip"
                filename = os.path.join(output_folder, nombre_archivo)
                fantasy_zip = zipfile.ZipFile(filename, 'w')

                wb = openpyxl.load_workbook(archivo)
                sheet = wb.worksheets[0]
                lista_ids = []
                lista_filas = []
                novalidos = 0
                # Recorrer las filas de la hoja
                for fila in range(1, sheet.max_row + 1):
                    if fila > 1:
                        codigo = sheet.cell(row=fila, column=1).value
                        if codigo:
                            codigo = codigo.strip()
                            if codigo_publicacion_valido(codigo, "ART"):
                                lista_ids.append(codigo.split(" ")[1].split("-")[0])
                            else:
                                lista_filas.append(fila)
                                novalidos += 1

                if novalidos > 0:
                    return JsonResponse({"result": "bad", "titulo": "Atención!!!", "mensaje": "{} {} {} que {} el código en formato incorrecto. Filas: {}".format("Existe" if novalidos == 1 else "Existen", novalidos, "registro" if novalidos == 1 else "registros", "tiene" if novalidos == 1 else "tienen", ", ".join(str(f) for f in lista_filas)), "showSwal": "True", "swalType": "warning"})

                anios = ArticuloInvestigacion.objects.filter(status=True, fechapublicacion__isnull=False).annotate(anio=ExtractYear('fechapublicacion')).values_list('anio', flat=True).order_by('-anio').distinct()

                for anio in anios:
                    print("Procesando año ", anio)
                    articulos_publicados = ArticuloInvestigacion.objects.filter(status=True, fechapublicacion__year=anio, pk__in=lista_ids).order_by('id')
                    articulos_no_publicados = ArticuloInvestigacion.objects.filter(status=True, fechapublicacion__isnull=True, fechaaprobacion__year=anio, pk__in=lista_ids).order_by('id')

                    total = articulos_publicados.count() + articulos_no_publicados.count()
                    rp = 0

                    if articulos_publicados or articulos_no_publicados:
                        print("Publicados: ", articulos_publicados.count())
                        print("No Publicados: ", articulos_no_publicados.count())
                        for articulo in articulos_publicados:
                            rp += 1
                            print("Procesando", rp, " de ", total)

                            titulo = articulo.nombre

                            palabras = titulo.split(" ")
                            titulo = "_".join(palabras[0:5])
                            titulo = remover_caracteres(titulo, caracteres_a_quitar)

                            titulo = elimina_tildes(remover_caracteres_tildes_unicode(remover_caracteres_especiales_unicode(titulo)))
                            carpetaarticulo = "ANIO_" + str(anio) + "/PUBLICADOS/ART_" + articulo.revista.codigoissn.strip().replace('-', '_') + "_" + str(articulo.id) + "_" + titulo

                            articulo_id = articulo.id
                            nombre = "ARTICULO_" + str(articulo_id).zfill(4)

                            # Agrego las evidencias a la carpeta del artículo
                            for evidencia in articulo.detalleevidencias_set.filter(status=True):
                                ext = evidencia.archivo.__str__()[evidencia.archivo.__str__().rfind("."):]

                                if evidencia.descripcion:
                                    nombreevidencia = evidencia.descripcion.upper()
                                    nombreevidencia = remover_caracteres(nombreevidencia, caracteres_a_quitar)
                                    nombreevidencia = elimina_tildes(remover_caracteres_tildes_unicode(remover_caracteres_especiales_unicode(nombreevidencia))).replace(" ", "_")
                                else:
                                    nombreevidencia = "EVIDENCIA" + random.randint(1, 10000).__str__()

                                if evidencia.evidencia.id == 1:
                                    if os.path.exists(SITE_STORAGE + evidencia.archivo.url):
                                        fantasy_zip.write(SITE_STORAGE + evidencia.archivo.url, carpetaarticulo + "/" + nombreevidencia + ext.lower())
                                elif evidencia.evidencia.id == 3:
                                    if os.path.exists(SITE_STORAGE + evidencia.archivo.url):
                                        fantasy_zip.write(SITE_STORAGE + evidencia.archivo.url, carpetaarticulo + "/" + nombreevidencia + ext.lower())

                            # Generar la ficha catalográfica del artículo
                            data['participantess'] = participantes = ParticipantesArticulos.objects.filter(articulo=articulo, status=True)
                            data['cantidad'] = participantes.count()
                            data['basesindexadas'] = ArticulosBaseIndexada.objects.filter(articulo=articulo, status=True)

                            nombrearchivoficha = 'fichacatalograficaart_' + str(articulo_id) + '_' + str(random.randint(1, 10000)) + '.pdf'
                            valida = convert_html_to_pdf(
                                'inv_articulos/fichacatalografica_articulo_pdf.html',
                                {'pagesize': 'A4', 'data': data},
                                nombrearchivoficha,
                                directorio
                            )

                            archivoficha = directorio + "/" + nombrearchivoficha
                            # Agrego el archivo de la ficha a la carpeta del artículo
                            fantasy_zip.write(archivoficha, carpetaarticulo + "/FICHA_CATALOGRAFICA.pdf")
                            # Borro el archivo de la ficha creado
                            ET.sleep(1)
                            os.remove(archivoficha)

                        for articulo in articulos_no_publicados:
                            rp += 1
                            print("Procesando", rp, " de ", total)

                            titulo = articulo.nombre

                            palabras = titulo.split(" ")
                            titulo = "_".join(palabras[0:5])
                            titulo = remover_caracteres(titulo, caracteres_a_quitar)

                            titulo = elimina_tildes(remover_caracteres_tildes_unicode(remover_caracteres_especiales_unicode(titulo)))
                            carpetaarticulo = "ANIO_" + str(anio) + "/ACEPTADOPUBLICACION/ART_" + articulo.revista.codigoissn.strip().replace('-', '_') + "_" + str(articulo.id) + "_" + titulo

                            articulo_id = articulo.id
                            nombre = "ARTICULO_" + str(articulo_id).zfill(4)

                            # Agrego las evidencias a la carpeta del artículo
                            for evidencia in articulo.detalleevidencias_set.filter(status=True):
                                ext = evidencia.archivo.__str__()[evidencia.archivo.__str__().rfind("."):]

                                if evidencia.descripcion:
                                    nombreevidencia = evidencia.descripcion.upper()
                                    nombreevidencia = remover_caracteres(nombreevidencia, caracteres_a_quitar)
                                    nombreevidencia = elimina_tildes(remover_caracteres_tildes_unicode(remover_caracteres_especiales_unicode(nombreevidencia))).replace(" ", "_")
                                else:
                                    nombreevidencia = "EVIDENCIA" + random.randint(1, 10000).__str__()

                                if evidencia.evidencia.id == 1:
                                    if os.path.exists(SITE_STORAGE + evidencia.archivo.url):
                                        fantasy_zip.write(SITE_STORAGE + evidencia.archivo.url, carpetaarticulo + "/" + nombreevidencia + ext.lower())
                                elif evidencia.evidencia.id == 3:
                                    if os.path.exists(SITE_STORAGE + evidencia.archivo.url):
                                        fantasy_zip.write(SITE_STORAGE + evidencia.archivo.url, carpetaarticulo + "/" + nombreevidencia + ext.lower())

                            # Generar la ficha catalográfica del artículo
                            data['participantess'] = participantes = ParticipantesArticulos.objects.filter(articulo=articulo, status=True)
                            data['cantidad'] = participantes.count()
                            data['basesindexadas'] = ArticulosBaseIndexada.objects.filter(articulo=articulo, status=True)

                            nombrearchivoficha = 'fichacatalograficaart_' + str(articulo_id) + '_' + str(random.randint(1, 10000)) + '.pdf'
                            valida = convert_html_to_pdf(
                                'inv_articulos/fichacatalografica_articulo_pdf.html',
                                {'pagesize': 'A4', 'data': data},
                                nombrearchivoficha,
                                directorio
                            )

                            archivoficha = directorio + "/" + nombrearchivoficha
                            # Agrego el archivo de la ficha a la carpeta del artículo
                            fantasy_zip.write(archivoficha, carpetaarticulo + "/FICHA_CATALOGRAFICA.pdf")
                            # Borro el archivo de la ficha creado
                            ET.sleep(1)
                            os.remove(archivoficha)

                fantasy_zip.close()
                ruta = "media/zipav/" + nombre_archivo
                return JsonResponse({'to': ruta})
            except Exception as ex:
                msg = ex.__str__()
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos. [%s]" % (msg)})

        elif action == 'delarticulo':
            try:
                # Consulto el artículo, participantes, bases indexadas y evidencias
                articulo = ArticuloInvestigacion.objects.get(pk=int(encrypt(request.POST['id'])))
                participantes = ParticipantesArticulos.objects.filter(articulo=articulo, matrizevidencia_id=3)
                bases = ArticulosBaseIndexada.objects.filter(articulo=articulo)
                evidencias = DetalleEvidencias.objects.filter(articulo=articulo)

                # Elimino los registros
                for participante in participantes:
                    participante.status = False
                    participante.save(request)
                for base in bases:
                    base.status = False
                    base.save(request)
                for evidencia in evidencias:
                    evidencia.status = False
                    evidencia.save(request)
                articulo.status = False
                articulo.save(request)

                log(u'%s eliminó artículo de investigación: %s' % (persona, articulo), request, "del")
                return JsonResponse({"error": False})
            except Exception as ex:
                msg = ex.__str__()
                transaction.set_rollback(True)
                return JsonResponse({"error": True, "message": u"Error al eliminar el registro. [%s]" % msg})

        elif action == 'aprobararticulo':
            try:
                articulo = ArticuloInvestigacion.objects.get(pk=int(encrypt(request.POST['id'])))

                # Si existe solicitud
                if articulo.solicitudpublicacion:
                    estadosolicitud = obtener_estado_solicitud(8, 5)  # APROBADO

                    solicitud = SolicitudPublicacion.objects.get(pk=articulo.solicitudpublicacion.id)
                    solicitud.aprobado = True
                    solicitud.estado = estadosolicitud
                    solicitud.save(request)

                articulo.aprobado = True
                articulo.save(request)

                integrantes = articulo.participantes()
                articulo.tipoaporte = 1 if integrantes.filter(tipoparticipanteins=1).count() >= 1 else 2
                articulo.save(request)

                # Si existe solicitud se debe notificar
                if articulo.solicitudpublicacion:
                    notificar_revision_solicitud_produccion_cientifica(solicitud)

                log(u'%s Aprobó artículo: %s' % (persona, articulo.nombre), request, "edit")
                return JsonResponse({"result": "ok", "titulo": "Proceso exitoso!!!", "mensaje": u"Registro actualizado con éxito", "showSwal": True})
            except Exception as ex:
                msg = ex.__str__()
                transaction.set_rollback(True)
                return JsonResponse({"error": True, "message": u"Error al aprobar el registro. [%s]" % msg})

        elif action == 'addgrupo':
            try:
                f = GrupoInvestigacionForm(request.POST)

                if f.is_valid():
                    if not GrupoInvestigacion.objects.filter(status=True, nombre=f.cleaned_data['nombre'].strip().upper()).exists():
                        integrantes = json.loads(request.POST['lista_items1'])

                        # Guardo el grupo
                        grupoinvestigacion = GrupoInvestigacion(
                            nombre=f.cleaned_data['nombre'],
                            descripcion=f.cleaned_data['descripcion'],
                            vigente=f.cleaned_data['vigente']
                        )
                        grupoinvestigacion.save(request)

                        # Guardo los integrantes del grupo de investigación
                        for integrante in integrantes:
                            integrantegrupo = GrupoInvestigacionIntegrante(
                                grupo=grupoinvestigacion,
                                funcion=integrante['idfuncion'],
                                persona_id=integrante['idpersona']
                            )
                            integrantegrupo.save(request)

                        log(u'%s agregó grupo de investigación %s' % (persona, grupoinvestigacion), request, "add")
                        return JsonResponse({"result": "ok", "titulo": "Proceso exitoso!!!", "mensaje": u"Registro guardado con éxito", "showSwal": True})
                    else:
                        return JsonResponse({"result": "bad", "titulo": "Atención!!!", "mensaje": u"El grupo de investigación ya ha sido ingresado", "showSwal": "True", "swalType": "warning"})
                else:
                    x = f.errors
                    raise NameError('Error en el formulario')
            except Exception as ex:
                msg = ex.__str__()
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "titulo": "Error", "mensaje": u"Error al guardar los datos. [%s]" % msg, "showSwal": "True", "swalType": "error"})

        elif action == 'editgrupo':
            try:
                f = GrupoInvestigacionForm(request.POST)

                if f.is_valid():
                    if not GrupoInvestigacion.objects.filter(status=True, nombre=f.cleaned_data['nombre'].strip().upper()).exclude(pk=int(encrypt(request.POST['id']))).exists():
                        # Consultar el grupo
                        grupoinvestigacion = GrupoInvestigacion.objects.get(pk=int(encrypt(request.POST['id'])))

                        integrantes = json.loads(request.POST['lista_items1'])
                        integranteseliminados = json.loads(request.POST['lista_items2']) if 'lista_items2' in request.POST else []

                        # Actualizo el grupo
                        grupoinvestigacion.nombre = f.cleaned_data['nombre']
                        grupoinvestigacion.descripcion = f.cleaned_data['descripcion']
                        grupoinvestigacion.vigente = f.cleaned_data['vigente']
                        grupoinvestigacion.save(request)

                        # Guardo los integrantes del grupo de investigación
                        for integrante in integrantes:
                            # Nuevo integrante
                            if int(integrante['idreg']) == 0:
                                integrantegrupo = GrupoInvestigacionIntegrante(
                                    grupo=grupoinvestigacion,
                                    funcion=integrante['idfuncion'],
                                    persona_id=integrante['idpersona']
                                )
                            else:
                                integrantegrupo = GrupoInvestigacionIntegrante.objects.get(pk=integrante['idreg'])
                                integrantegrupo.funcion = integrante['idfuncion']

                            integrantegrupo.save(request)

                        # Elimino los que se borraron del detalle
                        if integranteseliminados:
                            for integrante in integranteseliminados:
                                integrantegrupo = GrupoInvestigacionIntegrante.objects.get(pk=int(integrante['idreg']))
                                integrantegrupo.status = False
                                integrantegrupo.save(request)

                        log(u'%s editó grupo de investigación %s' % (persona, grupoinvestigacion), request, "edit")
                        return JsonResponse({"result": "ok", "titulo": "Proceso exitoso!!!", "mensaje": u"Registro actualizado con éxito", "showSwal": True})
                    else:
                        return JsonResponse({"result": "bad", "titulo": "Atención!!!", "mensaje": u"El grupo de investigación ya ha sido ingresado", "showSwal": "True", "swalType": "warning"})
                else:
                    x = f.errors
                    raise NameError('Error en el formulario')
            except Exception as ex:
                msg = ex.__str__()
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "titulo": "Error", "mensaje": u"Error al guardar los datos. [%s]" % msg, "showSwal": "True", "swalType": "error"})

        elif action == 'delgrupo':
            try:
                # Consulto el grupo de investigacion
                grupoinvestigacion = GrupoInvestigacion.objects.get(pk=int(encrypt(request.POST['id'])))

                # Verifico que se pueda eliminar
                if not grupoinvestigacion.puede_eliminar():
                    return JsonResponse({"result": "bad", "titulo": "Atención!!!", "mensaje": u"No se puede eliminar el Registro", "showSwal": "True", "swalType": "warning"})

                # Elimino el grupo de investigación
                grupoinvestigacion.status = False
                grupoinvestigacion.save(request)

                # Elimino los integrantes del grupo
                for integrante in grupoinvestigacion.integrantes():
                    integrante.status = False
                    integrante.save(request)

                log(u'%s eliminó grupo de investigación %s' % (persona, grupoinvestigacion), request, "del")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                msg = ex.__str__()
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "titulo": "Error", "mensaje": u"Error al eliminar el registro. [%s]" % msg, "showSwal": "True", "swalType": "error"})

        elif action == 'reasignarsolicitud':
            try:
                if not 'id' in request.POST or not 'tipodest' in request.POST:
                    return JsonResponse({"result": "bad", "titulo": "Error", "mensaje": u"Error al procesar la solicitud", "showSwal": "True", "swalType": "error"})

                # Consulto la solicitud de Grupo de Investigación
                grupoinvestigacion = GrupoInvestigacion.objects.get(pk=int(encrypt(request.POST['id'])))
                tipodest = request.POST['tipodest']

                # Si la persona a reasignar es al Coordinador de Investigación
                if tipodest == 'ci':
                    personareasignada = coordinador_investigacion()
                    # Obtener estado REASIGNADO A COORDINADOR
                    estado = obtener_estado_solicitud(19, 7)
                else:
                    personareasignada = tecnico_revisor_grupoinvestigacion()
                    # Obtener estado REASIGNADO A ANALISTA
                    estado = obtener_estado_solicitud(19, 8)

                # Actualizo la solicitud
                grupoinvestigacion.estado = estado
                grupoinvestigacion.save(request)

                # Guardar el recorrido
                recorrido = GrupoInvestigacionRecorrido(
                    grupo=grupoinvestigacion,
                    fecha=datetime.now().date(),
                    observacion=estado.observacion,
                    estado=estado
                )
                recorrido.save(request)

                # Notificar al solicitante
                notificar_grupo_investigacion(grupoinvestigacion, "NOTCOORD" if tipodest == 'ci' else "NOTANL")

                log(u'%s reasignó solicitud de propuesta de grupo de investigación para análisis: %s' % (persona, grupoinvestigacion), request, "edit")
                return JsonResponse({"result": "ok", "titulo": "Proceso exitoso!!!", "mensaje": u"Registro guardado con éxito", "showSwal": True})
            except Exception as ex:
                msg = ex.__str__()
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "titulo": "Error", "mensaje": u"Error al guardar los datos. [%s]" % msg, "showSwal": "True", "swalType": "error"})

        elif action == 'analizarsolicitud':
            try:
                if not 'id' in request.POST:
                    return JsonResponse({"result": "bad", "titulo": "Error", "mensaje": u"Error al procesar la solicitud", "showSwal": "True", "swalType": "error"})

                # Consulto la solicitud de Grupo de Investigación
                grupoinvestigacion = GrupoInvestigacion.objects.get(pk=int(encrypt(request.POST['id'])))
                estadoinicial = grupoinvestigacion.estado

                # Obtengo los valores del formulario
                requisitos = json.loads(request.POST['lista_items1'])
                valorestado = int(request.POST['estadosolicitud'])
                estado = obtener_estado_solicitud(19, valorestado)
                observacion = request.POST['observacion'].strip()

                # Actualizo la solicitud
                grupoinvestigacion.observacion = observacion
                grupoinvestigacion.estado = estado
                grupoinvestigacion.save(request)

                # Actualizar los requisitos que debe cumplir el director de grupo
                for requisito in requisitos:
                    integranterequisito = GrupoInvestigacionIntegranteRequisito.objects.get(pk=int(requisito['idreg']))
                    integranterequisito.cumpleanl = requisito['valor']
                    integranterequisito.save(request)

                # Si los estados son distintos
                if estadoinicial.valor != estado.valor:
                    # Guardar el recorrido
                    recorrido = GrupoInvestigacionRecorrido(
                        grupo=grupoinvestigacion,
                        fecha=datetime.now().date(),
                        observacion=observacion if observacion else estado.observacion,
                        estado=estado
                    )
                else:
                    recorrido = GrupoInvestigacionRecorrido.objects.filter(status=True, grupo=grupoinvestigacion, estado=estado).order_by('-id')[0]
                    recorrido.observacion = observacion if observacion else estado.observacion

                recorrido.save(request)

                # Notificar al solicitante
                notificar_grupo_investigacion(grupoinvestigacion, "VALSOL" if valorestado == 9 else "NOVANLSOL")

                log(u'%s analizó solicitud de propuesta para creación de grupo de investigación: %s' % (persona, grupoinvestigacion), request, "edit")
                return JsonResponse({"result": "ok", "titulo": "Proceso exitoso!!!", "mensaje": u"Registro guardado con éxito", "showSwal": True})
            except Exception as ex:
                msg = ex.__str__()
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "titulo": "Error", "mensaje": u"Error al guardar los datos. [%s]" % msg, "showSwal": "True", "swalType": "error"})

        elif action == 'devolverrequerimiento':
            try:
                if not 'id' in request.POST:
                    return JsonResponse({"result": "bad", "titulo": "Error", "mensaje": u"Error al procesar la solicitud", "showSwal": "True", "swalType": "error"})

                # Consulto la solicitud de Grupo de Investigación
                grupoinvestigacion = GrupoInvestigacion.objects.get(pk=int(encrypt(request.POST['id'])))
                observacion = ''
                resolucion = None

                tipodest = request.POST['tipodest']

                # Si la persona a reasignar es al Vicerrector de Investigación
                if tipodest == 'vi':
                    # Obtener estado DEVUELTO A VICERRECTOR
                    estado = obtener_estado_solicitud(19, 11)
                else:
                    # Obtener estado DEVUELTO A SOLICITANTE
                    estado = obtener_estado_solicitud(19, 12)

                # Actualizo campo aprobado por facultad
                if tipodest == 'sol':
                    # Consulto la resolucion
                    resolucion = grupoinvestigacion.resolucion_facultad()

                    # Actualizo la solicitud
                    grupoinvestigacion.aprobadocoord = False
                    observacion = grupoinvestigacion.observacion

                grupoinvestigacion.estado = estado
                grupoinvestigacion.save(request)

                # Si existe resolución actualizo campo vigente
                if resolucion:
                    resolucion.vigente = False
                    resolucion.save(request)

                # Guardar el recorrido
                recorrido = GrupoInvestigacionRecorrido(
                    grupo=grupoinvestigacion,
                    fecha=datetime.now().date(),
                    observacion=observacion if observacion else estado.observacion,
                    estado=estado
                )
                recorrido.save(request)

                # Notificar al solicitante
                notificar_grupo_investigacion(grupoinvestigacion, "DEVVICE" if tipodest == 'vi' else "DEVSOL")

                log(u'%s devolvió requerimiento de solicitud de propuesta de grupo de investigación: %s' % (persona, grupoinvestigacion), request, "edit")
                return JsonResponse({"result": "ok", "titulo": "Proceso exitoso!!!", "mensaje": u"Registro guardado con éxito", "showSwal": True})
            except Exception as ex:
                msg = ex.__str__()
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "titulo": "Error", "mensaje": u"Error al guardar los datos. [%s]" % msg, "showSwal": "True", "swalType": "error"})

        elif action == 'addinforme':
            try:
                if not 'id' in request.POST:
                    return JsonResponse({"result": "bad", "titulo": "Error", "mensaje": u"Error al enviar los datos", "showSwal": "True", "swalType": "error"})

                # Consulto la solciitud
                grupoinvestigacion = GrupoInvestigacion.objects.get(pk=int(encrypt(request.POST['id'])))

                objeto = "Solicitud de creación de " + grupoinvestigacion.nombre + ("(" + grupoinvestigacion.acronimo + ")" if grupoinvestigacion.acronimo else "")

                # Obtengo los campos del formulario
                motivaciontecnica = remover_atributo_style_html(request.POST['motivaciontecnica'])
                conclusion = remover_atributo_style_html(request.POST['conclusion'])
                recomendacion = remover_atributo_style_html(request.POST['recomendacion'])

                # Obtengo los valores del detalle de anexos
                nfilas_ca_evi = json.loads(request.POST['lista_items1'])  # Números de filas que tienen lleno el campo archivo en detalle de evidencias
                nfilas = request.POST.getlist('nfila_evidencia[]')  # Todos los número de filas del detalle de anexos
                descripciones = request.POST.getlist('descripcion_evidencia[]')  # Todas las descripciones
                archivos = request.FILES.getlist('archivo_evidencia[]')  # Todos los archivos
                fechasgenera = request.POST.getlist('fecha_genera[]')  # Todas las fechas de generación

                # Valido los archivos cargados en el detalle de anexos
                for nfila, archivo in zip(nfilas_ca_evi, archivos):
                    descripcionarchivo = 'Anexo'
                    resp = validar_archivo(descripcionarchivo, archivo, ['pdf'], '10MB')
                    if resp['estado'] != "OK":
                        return JsonResponse({"result": "bad", "titulo": "Atención!!!", "mensaje": resp["mensaje"] + " en la fila # " + str(nfila['cfila']), "showSwal": "True", "swalType": "warning"})

                # Obtengo estado INFORME DE PROPUESTA DE CREACIÓN DE GRUPO ELABORADO
                estado = obtener_estado_solicitud(19, 13)

                # Datos de personas que intervienen en el informe
                destinatario = vicerrector_investigacion_posgrado()
                cargodestinatario = destinatario.mi_cargo_actualadm().denominacionpuesto if destinatario.mi_cargo_actualadm() else destinatario.mi_cargo_actual().denominacionpuesto
                remitente = coordinador_investigacion()
                cargoremitente = remitente.mi_cargo_actualadm().denominacionpuesto if remitente.mi_cargo_actualadm() else remitente.mi_cargo_actual().denominacionpuesto
                elabora = persona
                cargoelabora = persona.mi_cargo_actualadm().denominacionpuesto if persona.mi_cargo_actualadm() else persona.mi_cargo_actual().denominacionpuesto
                verifica = experto_investigacion()
                cargoverifica = verifica.mi_cargo_actualadm().denominacionpuesto if verifica.mi_cargo_actualadm() else verifica.mi_cargo_actual().denominacionpuesto
                aprueba = remitente
                cargoaprueba = cargoremitente

                # Actualizar la solicitud
                grupoinvestigacion.informegen = True
                grupoinvestigacion.estado = estado
                grupoinvestigacion.save(request)

                # Obtener secuencia del informe
                secuencia = secuencia_informe_grupoinvestigacion()
                iniciales = persona.nombres[0] + persona.apellido1[0] + (persona.apellido2[0] if persona.apellido2 else '')
                numero = "ITI-VIP-GI-" + iniciales + "-" + str(secuencia).zfill(3) + "-" + str(datetime.now().year)

                # Crear el registro del informe
                informe = GrupoInvestigacionInforme(
                    grupo=grupoinvestigacion,
                    secuencia=secuencia,
                    fecha=datetime.now().date(),
                    numero=numero,
                    remitente=remitente,
                    cargoremitente=cargoremitente,
                    destinatario=destinatario,
                    cargodestinatario=cargodestinatario,
                    objeto=objeto,
                    antecedente='',
                    motivaciontecnica=motivaciontecnica,
                    conclusion=conclusion,
                    recomendacion=recomendacion,
                    elabora=elabora,
                    cargoelabora=cargoelabora,
                    verifica=verifica,
                    cargoverifica=cargoverifica,
                    aprueba=aprueba,
                    cargoaprueba=cargoaprueba,
                    estado=1
                )
                informe.save(request)

                # Guardo los anexos del informe
                for descripcion, archivo, fechagenera in zip(descripciones, archivos, fechasgenera):
                    fecha = datetime.strptime(fechagenera, '%Y-%m-%d').date()
                    pdf2ReaderEvi = PyPDF2.PdfFileReader(archivo)

                    archivoreg = archivo
                    archivoreg._name = generar_nombre("anexoinformegrupo", archivoreg._name)

                    anexoinforme = GrupoInvestigacionInformeAnexo(
                        informe=informe,
                        descripcion=descripcion.strip(),
                        fecha=fecha,
                        archivo=archivoreg,
                        numeropagina=pdf2ReaderEvi.numPages
                    )
                    anexoinforme.save(request)

                # Guardar el recorrido
                recorrido = GrupoInvestigacionRecorrido(
                    grupo=grupoinvestigacion,
                    fecha=datetime.now().date(),
                    observacion=estado.observacion,
                    estado=estado
                )
                recorrido.save(request)

                log(u'%s agregó informe técnico %s de creación de grupo de investigación a solicitud: %s' % (persona, informe.numero, grupoinvestigacion), request, "add")
                return JsonResponse({"result": "ok", "titulo": "Proceso exitoso!!!", "mensaje": u"Registro guardado con éxito", "showSwal": True, "id": request.POST['id']})
            except Exception as ex:
                msg = ex.__str__()
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "titulo": "Error", "mensaje": u"Error al guardar los datos. [%s]" % msg, "showSwal": "True", "swalType": "error"})

        elif action == 'editinforme':
            try:
                if not 'id' in request.POST:
                    return JsonResponse({"result": "bad", "titulo": "Error", "mensaje": u"Error al enviar los datos", "showSwal": "True", "swalType": "error"})

                # Consulto el informe
                informe = GrupoInvestigacionInforme.objects.get(pk=int(encrypt(request.POST['id'])))
                grupoinvestigacion = informe.grupo

                # Verifico que pueda editar
                if not grupoinvestigacion.puede_agregar_editar_informe():
                    return JsonResponse({"result": "bad", "titulo": "Atención!!!", "mensaje": u"No se puede actualizar el Registro porque ya ha sido editado por otra instancia", "showSwal": "True", "swalType": "warning"})

                # Obtengo los valores de los campos del formulario
                motivaciontecnica = remover_atributo_style_html(request.POST['motivaciontecnica'])
                conclusion = remover_atributo_style_html(request.POST['conclusion'])
                recomendacion = remover_atributo_style_html(request.POST['recomendacion'])

                # Obtengo los valores del detalle de anexos
                nfilas_ca_evi = json.loads(request.POST['lista_items1'])  # Números de filas que tienen lleno el campo archivo en detalle de evidencias
                nfilas = request.POST.getlist('nfila_evidencia[]')  # Todos los número de filas del detalle de anexos
                idsanexos = request.POST.getlist('idregistro[]')  # Todos los ids de detalle de anexos
                descripciones = request.POST.getlist('descripcion_evidencia[]')  # Todas las descripciones
                archivos = request.FILES.getlist('archivo_evidencia[]')  # Todos los archivos
                fechasgenera = request.POST.getlist('fecha_genera[]')  # Todas las fechas de generación
                anexosseliminados = json.loads(request.POST['lista_items2']) if 'lista_items2' in request.POST else []

                # Valido los archivos cargados en el detalle de anexos
                for nfila, archivo in zip(nfilas_ca_evi, archivos):
                    descripcionarchivo = 'Anexo'
                    resp = validar_archivo(descripcionarchivo, archivo, ['pdf'], '10MB')
                    if resp['estado'] != "OK":
                        return JsonResponse({"result": "bad", "titulo": "Atención!!!", "mensaje": resp["mensaje"] + " en la fila # " + str(nfila['cfila']), "showSwal": "True", "swalType": "warning"})

                # En caso que el estado sea NOVEDAD o se haya ENVIADO al EXPERTO, se debe actualizar los estados de la postulacion
                actualizarpostulacion = informe.estado == 3 or grupoinvestigacion.estado.valor == 14

                if actualizarpostulacion:
                    # Obtengo estado INFORME DE PROPUESTA DE CREACIÓN DE GRUPO ELABORADO
                    estado = obtener_estado_solicitud(19, 13)

                    # Actualizar la solicitud
                    grupoinvestigacion.observacion = ""
                    grupoinvestigacion.estado = estado
                    grupoinvestigacion.save(request)

                # Actualizo el registro del informe
                informe.motivaciontecnica = motivaciontecnica
                informe.conclusion = conclusion
                informe.recomendacion = recomendacion
                informe.archivo = None
                informe.archivofirmado = None
                informe.impreso = False
                informe.firmaelabora = False
                informe.firmaverifica = False
                informe.firmaaprueba = False
                informe.observacion = ""
                informe.estado = 1
                informe.save(request)

                # Guardo los anexos del informe
                for idanexo, nfila, descripcion, fechagenera in zip(idsanexos, nfilas, descripciones, fechasgenera):
                    fecha = datetime.strptime(fechagenera, '%Y-%m-%d').date()

                    # Si es registro nuevo
                    if int(idanexo) == 0:
                        anexoinforme = GrupoInvestigacionInformeAnexo(
                            informe=informe,
                            descripcion=descripcion.strip(),
                            fecha=fecha
                        )
                    else:
                        anexoinforme = GrupoInvestigacionInformeAnexo.objects.get(pk=idanexo)
                        anexoinforme.descripcion = descripcion.strip()
                        anexoinforme.fecha = fecha

                    anexoinforme.save(request)

                    # Guardo el archivo del anexo
                    for nfilaarchi, archivo in zip(nfilas_ca_evi, archivos):
                        # Si la fila de la descripcion es igual a la fila que contiene archivo
                        if int(nfilaarchi['nfila']) == int(nfila):
                            pdf2ReaderEvi = PyPDF2.PdfFileReader(archivo)

                            archivoreg = archivo
                            archivoreg._name = generar_nombre("anexoinformegrupo", archivoreg._name)

                            # actualizo campo archivo del registro creado
                            anexoinforme.archivo = archivoreg
                            anexoinforme.numeropagina = pdf2ReaderEvi.numPages
                            anexoinforme.save(request)
                            break

                # Elimino los anexos que se borraron del detalle
                if anexosseliminados:
                    for anexo in anexosseliminados:
                        anexoinforme = GrupoInvestigacionInformeAnexo.objects.get(pk=anexo['idreg'])
                        anexoinforme.status = False
                        anexoinforme.save(request)

                if actualizarpostulacion:
                    # Guardar el recorrido
                    recorrido = GrupoInvestigacionRecorrido(
                        grupo=grupoinvestigacion,
                        fecha=datetime.now().date(),
                        observacion=estado.observacion,
                        estado=estado
                    )
                    recorrido.save(request)

                log(u'%s editó informe técnico %s de creación de grupo de investigación a solicitud: %s' % (persona, informe.numero, grupoinvestigacion), request, "edit")
                return JsonResponse({"result": "ok", "titulo": "Proceso exitoso!!!", "mensaje": u"Registro actualizado con éxito", "showSwal": True, "id": request.POST['ids']})
            except Exception as ex:
                msg = ex.__str__()
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "titulo": "Error", "mensaje": u"Error al guardar los datos. [%s]" % msg, "showSwal": "True", "swalType": "error"})

        elif action == 'generarinforme':
            try:
                data = {}

                informe = GrupoInvestigacionInforme.objects.get(pk=int(encrypt(request.POST['id'])))

                data['informe'] = informe
                data['grupo'] = grupoinvestigacion = informe.grupo
                data['resolucion'] = resolucion = grupoinvestigacion.resolucion_facultad()
                data['fecharesolucion'] = str(resolucion.fecha.day) + " de " + MESES_CHOICES[resolucion.fecha.month - 1][1].capitalize() + " del " + str(resolucion.fecha.year)
                data['decano'] = responsable_coordinacion(grupoinvestigacion.periodo, grupoinvestigacion.coordinacion).persona
                data['coordinador'] = informe.remitente

                anexos = []
                numero = 0
                anexos_documentos = grupoinvestigacion.resolucion_memorando()
                anexos_adicionales = informe.anexos()

                for anexo in anexos_documentos:
                    numero += 1
                    anexos.append({
                        "numero": numero,
                        "descripcion": anexo["descripcion"],
                        "fecha": anexo["fecha"],
                        "numeropagina": anexo["numeropagina"],
                        "archivo": anexo["archivo"]
                    })

                for anexo in anexos_adicionales:
                    numero += 1
                    anexos.append({
                        "numero": numero,
                        "descripcion": anexo.descripcion,
                        "fecha": anexo.fecha,
                        "numeropagina": anexo.numeropagina,
                        "archivo": anexo.archivo
                    })

                data['anexos'] = anexos

                # Creacion de los archivos por separado
                directorio = SITE_STORAGE + '/media/certificadoedocente'
                try:
                    os.stat(directorio)
                except:
                    os.mkdir(directorio)

                # Archivo de la parte 1 del informe
                nombrearchivo = 'informeparte1_' + str(informe.id) + '_' + str(random.randint(1, 10000)) + '.pdf'
                valida = convert_html_to_pdf(
                    'ges_grupoinvestigacion/informetecnicopdf.html',
                    {'pagesize': 'A4', 'data': data},
                    nombrearchivo,
                    directorio
                )

                if not valida:
                    return JsonResponse({"result": "bad", "titulo": "Error", "mensaje": u"Error al generar el documento de la parte 1 del informe. [%s]", "showSwal": "True", "swalType": "error"})

                archivo1 = directorio + "/" + nombrearchivo

                # Leer los archivos
                pdf1Reader = PyPDF2.PdfFileReader(archivo1)

                # Crea un nuevo objeto PdfFileWriter que representa un documento PDF en blanco
                pdfWriter = PyPDF2.PdfFileWriter()

                # Recorre todas las páginas del documento 1
                for pageNum in range(pdf1Reader.numPages):
                    pageObj = pdf1Reader.getPage(pageNum)
                    pdfWriter.addPage(pageObj)

                # Recorro el detalle de requisitos
                for evidencia in anexos:
                    archivoevidencia = SITE_STORAGE + evidencia["archivo"].url  # Archivo pdf cargado
                    pdf2ReaderEvi = PyPDF2.PdfFileReader(archivoevidencia)

                    # Recorre todas las páginas del documento de evidencia: pdf2ReaderEvi
                    for pageNum in range(pdf2ReaderEvi.numPages):
                        pageObj = pdf2ReaderEvi.getPage(pageNum)
                        pdfWriter.addPage(pageObj)

                # Luego de haberse copiado todas las paginas de todos los documentos, se las escribe en el documento en blanco
                fecha = datetime.now().date()
                hora = datetime.now().time()
                nombrearchivoresultado = generar_nombre('informetecnicogrupo', 'informetecnicogrupo.pdf')
                pdfOutputFile = open(directorio + "/" + nombrearchivoresultado, "wb")
                pdfWriter.write(pdfOutputFile)

                # Borro los documento individuales creados a exepción del archivo del proyecto cargado por el docente
                os.remove(archivo1)

                pdfOutputFile.close()

                archivo = SITE_STORAGE + '/media/certificadoedocente/' + nombrearchivoresultado

                # Aperturo el archivo generado
                with open(archivo, 'rb') as f:
                    data = f.read()

                buffer = io.BytesIO()
                buffer.write(data)
                pdfcopia = buffer.getvalue()
                buffer.seek(0)
                buffer.close()

                # Extraigo el contenido
                archivocopiado = ContentFile(pdfcopia)
                archivocopiado.name = nombrearchivoresultado

                # Actualizo el informe
                informe.impreso = True
                informe.archivo = archivocopiado
                informe.save(request)

                # Borro el informe creado de manera general, no la del registro
                os.remove(archivo)

                return JsonResponse({"result": "ok", "idi": encrypt(informe.id), "id": encrypt(grupoinvestigacion.id)})
                # return JsonResponse({"result": "ok", "documento": informe.archivo.url, "id": encrypt(grupoinvestigacion.id)})
                # return JsonResponse({"result": "ok", "documento": archivo})
            except Exception as ex:
                msg = ex.__str__()
                return JsonResponse({"result": "bad", "titulo": "Error", "mensaje": u"Error al generar el documento del informe. [%s]" % msg, "showSwal": "True", "swalType": "error"})

        elif action == 'firmarinforme':
            try:
                if 'iddoc' not in request.POST:
                    return JsonResponse({"result": "bad", "titulo": "Error", "mensaje": u"Error al procesar la solicitud", "showSwal": "True", "swalType": "error"})

                tipofirma = request.POST['tipofirma']
                archivofirma = request.FILES['archivofirma']
                clavefirma = request.POST['cfirma']
                descripcionarchivo = 'Archivo de la firma electrónica'

                # Validar el archivo
                resp = validar_archivo(descripcionarchivo, archivofirma, ['P12', 'PFX'], '4MB')
                if resp['estado'] != "OK":
                    return JsonResponse({"result": "bad", "titulo": "Atención!!!", "mensaje": resp["mensaje"], "showSwal": "True", "swalType": "warning"})

                # Consulto el informe
                informe = GrupoInvestigacionInforme.objects.get(pk=int(encrypt(request.POST['iddoc'])))
                grupoinvestigacion = informe.grupo

                # Verifico si puede firmar el informe
                if tipofirma == 'ELA':
                    if not grupoinvestigacion.puede_firmar_informe():
                        return JsonResponse({"result": "bad", "titulo": "Atención!!!", "mensaje": u"No se puede editar el Registro porque ya ha sido revisado por otra instancia", "showSwal": "True", "swalType": "warning"})
                elif tipofirma == 'VER':
                    if not grupoinvestigacion.puede_firmar_informe_experto():
                        return JsonResponse({"result": "bad", "titulo": "Atención!!!", "mensaje": u"No se puede editar el Registro porque ya ha sido revisado por otra instancia", "showSwal": "True", "swalType": "warning"})
                else:
                    if not grupoinvestigacion.puede_firmar_informe_coordinador():
                        return JsonResponse({"result": "bad", "titulo": "Atención!!!", "mensaje": u"No se puede editar el Registro porque ya ha sido revisado por otra instancia", "showSwal": "True", "swalType": "warning"})

                # Obtengo el archivo del informe
                archivoinforme = informe.archivo if not informe.archivofirmado else informe.archivofirmado
                rutapdfarchivo = SITE_STORAGE + archivoinforme.url

                if tipofirma == 'ELA':
                    textoabuscar = informe.elabora.nombre_completo_inverso()
                    textofirma = 'Elaborado por:'
                    ocurrencia = 1
                elif tipofirma == 'VER':
                    textoabuscar = informe.verifica.nombre_completo_inverso()
                    textofirma = 'Revisado por:'
                    ocurrencia = 1
                else:
                    textoabuscar = informe.aprueba.nombre_completo_inverso()
                    textofirma = 'Aprobado por:'
                    ocurrencia = 1

                vecescontrado = 0
                documento = fitz.open(rutapdfarchivo)
                numpaginafirma = int(documento.page_count) - 1

                # Busca la página donde se encuentran ubicados los textos: Elaboraro por, Verificado por y Aprobado por
                words_dict = {}
                encontrado = False
                for page_number, page in enumerate(documento):
                    words = page.get_text("blocks")
                    words_dict[0] = words

                    for cadena in words_dict[0]:
                        linea = cadena[4].replace("\n", " ")
                        if linea:
                            linea = linea.strip()

                        if textofirma in linea:
                            numpaginafirma = page_number
                            encontrado = True
                            break

                    if encontrado:
                        break

                # Obtener coordenadas para ubicar al firma dependiendo de donde se encuentra ubicado el nombre de la persona
                saltolinea = False
                words_dict = {}
                for page_number, page in enumerate(documento):
                    if page_number == numpaginafirma:
                        words = page.get_text("blocks")
                        words_dict[0] = words

                valor = None
                for cadena in words_dict[0]:
                    linea = cadena[4].replace("\n", " ")
                    if linea:
                        linea = linea.strip()

                    if textoabuscar in linea:
                        valor = cadena

                        vecescontrado += 1
                        if vecescontrado == ocurrencia:
                            break

                if valor:
                    y = 5000 - int(valor[3]) - 4120
                else:
                    y = 0

                # x = 87  # izq
                # x = 230  # cent
                x = 350  # der

                # Si no existe el nombre de la persona en el documento
                if y == 0:
                    return JsonResponse({"result": "bad", "titulo": "Error", "mensaje": u"No se encuentra el nombre del firmante en el documento", "showSwal": "True", "swalType": "warning"})

                # Obtener extensión y leer archivo de la firma
                extfirma = os.path.splitext(archivofirma.name)[1][1:]
                bytesfirma = archivofirma.read()

                # Firma del documento
                datau = JavaFirmaEc(
                    archivo_a_firmar=archivoinforme,
                    archivo_certificado=bytesfirma,
                    extension_certificado=extfirma,
                    password_certificado=clavefirma,
                    page=numpaginafirma,
                    reason='',
                    lx=x,
                    ly=y
                ).sign_and_get_content_bytes()

                generar_archivo_firmado = io.BytesIO()
                generar_archivo_firmado.write(datau)
                generar_archivo_firmado.seek(0)

                nombrearchivofirmado = generar_nombre('informetecnicogrupofirmado', 'informetecnicogrupofirmado.pdf')
                objarchivo = DjangoFile(generar_archivo_firmado, nombrearchivofirmado)

                informe.archivofirmado = objarchivo

                if tipofirma == 'ELA':
                    informe.firmaelabora = True
                elif tipofirma == 'VER':
                    informe.firmaverifica = True
                    informe.estado = 2
                else:
                    informe.firmaaprueba = True
                    informe.estado = 4

                informe.save(request)

                if tipofirma == 'ELA':
                    if grupoinvestigacion.estado.valor == 13:
                        # Obtengo estado INFORME ENVIADO AL EXPERTO DE INVESTIGACIÓN
                        estado = obtener_estado_solicitud(19, 14)

                        # Actualizo la solicitud
                        grupoinvestigacion.estado = estado
                        grupoinvestigacion.save(request)

                        # Guardar el recorrido
                        recorrido = GrupoInvestigacionRecorrido(
                            grupo=grupoinvestigacion,
                            fecha=datetime.now().date(),
                            observacion=estado.observacion,
                            estado=estado
                        )
                        recorrido.save(request)

                    # Notificar por email al Experto de Investigacion
                    notificar_grupo_investigacion(grupoinvestigacion, "INFOELA")
                elif tipofirma == 'VER':
                    if grupoinvestigacion.estado.valor == 15:
                        # Obtengo estado INFORME ENVIADO AL COORDINADOR DE INVESTIGACIÓN
                        estado = obtener_estado_solicitud(19, 16)

                        # Actualizo la solicitud
                        grupoinvestigacion.estado = estado
                        grupoinvestigacion.save(request)

                        # Guardar el recorrido
                        recorrido = GrupoInvestigacionRecorrido(
                            grupo=grupoinvestigacion,
                            fecha=datetime.now().date(),
                            observacion=estado.observacion,
                            estado=estado
                        )
                        recorrido.save(request)

                    # Notificar por email al Coordinador de Investigacoón
                    notificar_grupo_investigacion(grupoinvestigacion, "INFOVAL")
                else:
                    if grupoinvestigacion.estado.valor == 16:
                        # Obtengo estado INFORME APROBADO POR COORDINADOR DE INVESTIGACIÓN
                        estado = obtener_estado_solicitud(19, 18)

                        # Actualizo la solicitud
                        grupoinvestigacion.estado = estado
                        grupoinvestigacion.save(request)

                        # Guardar el recorrido
                        recorrido = GrupoInvestigacionRecorrido(
                            grupo=grupoinvestigacion,
                            fecha=datetime.now().date(),
                            observacion=estado.observacion,
                            estado=estado
                        )
                        recorrido.save(request)

                        # Obtengo estado INFORME REMITIDO AL VICERRECTOR
                        estado = obtener_estado_solicitud(19, 19)

                        # Actualizo la solicitud
                        grupoinvestigacion.estado = estado
                        grupoinvestigacion.save(request)

                        # Guardar el recorrido
                        recorrido = GrupoInvestigacionRecorrido(
                            grupo=grupoinvestigacion,
                            fecha=datetime.now().date(),
                            observacion=estado.observacion,
                            estado=estado
                        )
                        recorrido.save(request)

                    # Notificar por email al Vicerrector de Investigación y Posgrado
                    notificar_grupo_investigacion(grupoinvestigacion, "INFOAPR")

                log(u'%s firmó informe técnico de creación de grupo de investigación: %s' % (persona, informe), request, "edit")
                return JsonResponse({"result": "ok", "titulo": "Proceso exitoso!!!", "mensaje": u"Documento firmado con éxito", "showSwal": True, "idi": encrypt(informe.id), "id": encrypt(grupoinvestigacion.id)})
            except Exception as ex:
                msg = ex.__str__()
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "titulo": "Error", "mensaje": u"Error al firmar el documento. [%s]" % msg, "showSwal": "True", "swalType": "error"})

        elif action == 'validarinforme':
            try:
                if 'id' not in request.POST:
                    return JsonResponse({"result": "bad", "titulo": "Error", "mensaje": u"Error al procesar la solicitud", "showSwal": "True", "swalType": "error"})

                # Consulto el informe técnico
                informe = GrupoInvestigacionInforme.objects.get(pk=int(encrypt(request.POST['id'])))
                grupoinvestigacion = informe.grupo
                estadoinicial = grupoinvestigacion.estado

                # Si no se puede validar el informe
                if not grupoinvestigacion.puede_validar_informe():
                    return JsonResponse({"result": "bad", "titulo": "Atención!!!", "mensaje": u"No se puede editar el Registro porque ya ha sido revisado por otra instancia", "showSwal": "True", "swalType": "warning"})

                estado = int(request.POST['estado'])
                observacion = request.POST['observacion'].strip().upper() if 'observacion' in request.POST else ''

                # Si estado es VALIDADO
                if estado == 2:
                    # Obtengo estado INFORME VALIDADO POR EXPERTO DE INVESTIGACIÓN
                    estadosolicitud = obtener_estado_solicitud(19, 15)
                else:
                    # Obtengo estado NOVEDADES EN LA VALIDACIÓN DEL INFORME
                    estadosolicitud = obtener_estado_solicitud(19, 17)

                # Actualizo la solicitud
                grupoinvestigacion.observacion = observacion
                grupoinvestigacion.estado = estadosolicitud
                grupoinvestigacion.save(request)

                # Actualizo el informe
                informe.observacion = observacion
                informe.estado = estado
                informe.save(request)

                # Si los estados son distintos
                if estadoinicial.valor != estadosolicitud.valor:
                    # Guardar el recorrido
                    recorrido = GrupoInvestigacionRecorrido(
                        grupo=grupoinvestigacion,
                        fecha=datetime.now().date(),
                        observacion=observacion if observacion else estadosolicitud.observacion,
                        estado=estadosolicitud
                    )
                else:
                    recorrido = GrupoInvestigacionRecorrido.objects.filter(status=True, grupo=grupoinvestigacion, estado=estadosolicitud).order_by('-id')[0]
                    recorrido.observacion = observacion

                recorrido.save(request)

                # Si tiene novedades se notifica al que elaboró el informe
                if estadosolicitud.valor == 17:
                    # Notificar por email al Analista de Investigacion
                    notificar_grupo_investigacion(grupoinvestigacion, "INFONOV")

                return JsonResponse({"result": "ok", "titulo": "Proceso exitoso!!!", "mensaje": u"Registro guardado con éxito", "showSwal": True})
            except Exception as ex:
                msg = ex.__str__()
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "titulo": "Error", "mensaje": u"Error al guardar los datos. [%s]" % msg, "showSwal": "True", "swalType": "error"})

        elif action == 'descargarinformes':
            try:
                caracteres_a_quitar = "!\"\#$%&()=?¡'¿{}[],.-+*/|°^~:;"

                output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media', 'zipav'))

                response = HttpResponse(content_type='application/zip')
                response['Content-Disposition'] = 'attachment; filename=informes_gruposinvestigacion_' + random.randint(1, 10000).__str__() + '.zip'

                nombre_archivo = generar_nombre('informesgruposinvestigacion', 'informesgruposinvestigacion.zip')
                filename = os.path.join(output_folder, nombre_archivo)
                fantasy_zip = zipfile.ZipFile(filename, 'w')

                # Consulto los registros de informes aprobados
                informes = GrupoInvestigacionInforme.objects.filter(status=True, firmaaprueba=True).order_by('secuencia')

                for informe in informes:
                    if informe.archivofirmado:
                        titulo = informe.grupo.nombre.upper()
                        palabras = titulo.split(" ")
                        titulo = "_".join(palabras[0:5])
                        titulo = remover_caracteres(titulo, caracteres_a_quitar)

                        titulo = elimina_tildes(remover_caracteres_tildes_unicode(remover_caracteres_especiales_unicode(titulo)))
                        nombrearchivo = informe.numero.replace("-", "_") + "_" + titulo

                        ext = informe.archivofirmado.__str__()[informe.archivofirmado.__str__().rfind("."):]

                        if os.path.exists(SITE_STORAGE + informe.archivofirmado.url):
                            fantasy_zip.write(SITE_STORAGE + informe.archivofirmado.url, nombrearchivo + ext.lower())

                fantasy_zip.close()

                ruta_zip = "media/zipav/" + nombre_archivo

                return JsonResponse({'result': 'ok', 'archivo': ruta_zip})
            except Exception as ex:
                msg = ex.__str__()
                return JsonResponse({"result": "bad", "titulo": "Error", "mensaje": u"Error al descargar los informes. Detalle: [%s]" % msg, "showSwal": "True", "swalType": "error"})

        elif action == 'marcarremitidocga':
            try:
                # Consultar los informes que no han sido remitidos a CGA
                informes = GrupoInvestigacionInforme.objects.filter(status=True, firmaaprueba=True, remitidocga=False).order_by('id')

                # Obtengo estado INFORME REMITIDO PARA TRATAMIENTO A CGA
                estado = obtener_estado_solicitud(19, 20)

                for informe in informes:
                    # Actualizo el informe
                    informe.remitidocga = True
                    informe.save(request)

                    # Obtengo el grupo de investigación para poder actualizar estados
                    grupoinvestigacion = informe.grupo

                    grupoinvestigacion.estado = estado
                    grupoinvestigacion.save(request)

                    # Guardar el recorrido
                    recorrido = GrupoInvestigacionRecorrido(
                        grupo=grupoinvestigacion,
                        fecha=datetime.now().date(),
                        observacion=estado.observacion,
                        estado=estado
                    )
                    recorrido.save(request)

                    log(u'%s marcó como remitido el informe técnico %s de creación de grupo de investigación a solicitud: %s' % (persona, informe.numero, grupoinvestigacion), request, "edit")

                return JsonResponse({"result": "ok", "titulo": "Proceso exitoso!!!", "mensaje": u"Registros actualizados con éxito", "showSwal": True})
            except Exception as ex:
                msg = ex.__str__()
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "titulo": "Error", "mensaje": u"Error al guardar los datos. [%s]" % msg, "showSwal": "True", "swalType": "error"})

        elif action == 'aprobacionocs':
            try:
                if 'id' not in request.POST:
                    return JsonResponse({"result": "bad", "titulo": "Error", "mensaje": u"Error al procesar la solicitud", "showSwal": "True", "swalType": "error"})

                # Consulto la solicitud de Grupo de Investigación
                grupoinvestigacion = GrupoInvestigacion.objects.get(pk=int(encrypt(request.POST['id'])))
                estadoinicial = grupoinvestigacion.estado
                archivoresolucion = None

                # Verifico que se pueda editar
                if not grupoinvestigacion.puede_subir_resolucion_ocs():
                    return JsonResponse({"result": "bad", "titulo": "Atención!!!", "mensaje": u"No se puede editar el Registro porque ya ha sido revisado por otra instancia", "showSwal": "True", "swalType": "warning"})

                if 'archivoresolucion' in request.FILES:
                    archivoresolucion = request.FILES['archivoresolucion']
                    descripcionarchivo = 'Archivo Resolución OCS'

                    # Validar el archivo
                    resp = validar_archivo(descripcionarchivo, archivoresolucion, ['PDF'], '4MB')
                    if resp['estado'] != "OK":
                        return JsonResponse({"result": "bad", "titulo": "Atención!!!", "mensaje": resp["mensaje"], "showSwal": "True", "swalType": "warning"})

                    archivoresolucion = request.FILES['archivoresolucion']
                    archivoresolucion._name = generar_nombre("resolucionocs", archivoresolucion._name)

                # Obtengo los valores del formulario
                valorestado = int(request.POST['estadosolicitud'])
                estado = obtener_estado_solicitud(19, valorestado)
                numeroresolucion = request.POST['numeroresolucion'].strip().upper()
                fecharesolucion = datetime.strptime(request.POST['fecharesolucion'], '%Y-%m-%d').date()
                resuelve = ''

                # Fecha de resolución debe ser mayor o igual a fecha resolución consejo
                if fecharesolucion < grupoinvestigacion.resolucion_facultad().fecha:
                    return JsonResponse({"result": "bad", "titulo": "Atención!!!", "mensaje": u"La fecha de resolución OCS debe ser mayor o igual a la fecha de resolución de Consejo Directivo de Facultad", "showSwal": "True", "swalType": "warning"})

                # Actualizo la solicitud
                grupoinvestigacion.solicitudvigente = False
                grupoinvestigacion.vigente = True
                grupoinvestigacion.aprobadoocs = True
                grupoinvestigacion.estado = estado
                grupoinvestigacion.save(request)

                # Consulto la resolución de ocs
                resolucion = grupoinvestigacion.resolucion_ocs()
                if not resolucion:
                    pdf2ReaderEvi = PyPDF2.PdfFileReader(archivoresolucion)
                    resolucion = GrupoInvestigacionResolucion(
                        grupo=grupoinvestigacion,
                        tipo=2,
                        numero=numeroresolucion,
                        fecha=fecharesolucion,
                        resuelve=resuelve,
                        archivo=archivoresolucion,
                        numeropagina=pdf2ReaderEvi.numPages,
                        vigente=True
                    )
                else:
                    resolucion.numero = numeroresolucion
                    resolucion.fecha = fecharesolucion
                    resolucion.resuelve = resuelve
                    if archivoresolucion:
                        pdf2ReaderEvi = PyPDF2.PdfFileReader(archivoresolucion)
                        resolucion.archivo = archivoresolucion
                        resolucion.numeropagina = pdf2ReaderEvi.numPages

                resolucion.save(request)

                # Si el estado original es REQUERIMIENTO REMITIDO PARA TRATAMIENTO A CGA, notificar por email
                if estadoinicial.valor == 20:
                    # Guardar el recorrido
                    recorrido = GrupoInvestigacionRecorrido(
                        grupo=grupoinvestigacion,
                        fecha=datetime.now().date(),
                        observacion=estado.observacion,
                        estado=estado
                    )
                    recorrido.save(request)

                    # Notificar al solicitante
                    notificar_grupo_investigacion(grupoinvestigacion, "APROCS")

                    # Notificar al vicerrector de investigación
                    notificar_grupo_investigacion(grupoinvestigacion, "NOTVICEOCS")

                    # Notificar al coordinador de investigación
                    notificar_grupo_investigacion(grupoinvestigacion, "NOTCOORDOCS")

                log(u'%s registró aprobación de OCS para la solicitud de propuesta para creación de grupo de investigación: %s' % (persona, grupoinvestigacion), request, "edit")
                return JsonResponse({"result": "ok", "titulo": "Proceso exitoso!!!", "mensaje": u"Registro guardado con éxito", "showSwal": True})
            except Exception as ex:
                msg = ex.__str__()
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "titulo": "Error", "mensaje": u"Error al guardar los datos. [%s]" % msg, "showSwal": "True", "swalType": "error"})


        return JsonResponse({"result": "bad", "titulo": "Error", "mensaje": u"Solicitud incorrecta.", "showSwal": "True", "swalType": "error"})
    else:
        if 'action' in request.GET:
            data['action'] = action = request.GET['action']

            if action == 'reportegeneralarticulos':
                try:
                    output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media', 'postgrado'))
                    tipo = request.GET['tipo']
                    participantes = request.GET['participantes']

                    if tipo == 'gen':
                        nombrearchivo = "LISTA_ARTICULOS_" + datetime.now().strftime('%Y%m%d_%H%M%S') + ".xlsx"
                    else:
                        nombrearchivo = "LISTA_ARTICULOS_PARTICIPANTES_" + datetime.now().strftime('%Y%m%d_%H%M%S') + ".xlsx"

                    # Crea un nuevo archivo de excel y le agrega una hoja
                    workbook = xlsxwriter.Workbook(output_folder + '/' + nombrearchivo)
                    ws = workbook.add_worksheet("Listado")

                    fcabeceracolumna = workbook.add_format(FORMATOS_CELDAS_EXCEL["cabeceracolumna"])
                    fceldageneral = workbook.add_format(FORMATOS_CELDAS_EXCEL["celdageneral"])
                    fceldageneralcent = workbook.add_format(FORMATOS_CELDAS_EXCEL["celdageneralcent"])
                    ftitulo1 = workbook.add_format(FORMATOS_CELDAS_EXCEL["titulo1"])
                    fceldafechaDMA = workbook.add_format(FORMATOS_CELDAS_EXCEL["celdafechaDMA"])

                    ws.merge_range(0, 0, 0, 35, 'UNIVERSIDAD ESTATAL DE MILAGRO', ftitulo1)
                    ws.merge_range(1, 0, 1, 35, 'VICERRECTORADO DE INVESTIGACIÓN Y POSGRADO', ftitulo1)
                    ws.merge_range(2, 0, 2, 35, 'COORDINACIÓN DE INVESTIGACIÓN', ftitulo1)
                    ws.merge_range(3, 0, 3, 35, 'LISTADO GENERAL DE ARTÍCULOS' if tipo == 'gen' else 'LISTADO GENERAL DE ARTÍCULOS POR PARTICIPANTES', ftitulo1)

                    columns = [
                        (u"TIPO DE ARTÍCULOS", 11),
                        (u"CÓDIGO", 17),
                        (u"ARTÍCULO", 39),
                        (u"VOL.", 8),
                        (u"NUM.", 8),
                        (u"PAG.", 8),
                        (u"CÓDIGO ISSN", 16),
                        (u"BASE INDEXADA", 8),
                        (u"NOMBRE BASE INDEXADA", 39),
                        (u"REVISTA", 39),
                        (u"TIPO", 17),
                        (u"SJR", 8),
                        (u"FECHA PUBLICACIÓN", 11),
                        (u"ÁREA DE CONOCIMIENTO", 39),
                        (u"SUB-ÁREA DE CONOCIMIENTO", 39),
                        (u"SUB-ÁREA ESPECÍFICA", 39),
                        (u"LÍNEA INVESTIGACIÓN", 39),
                        (u"SUB-LÍNEA INVESTIGACIÓN", 39),
                        (u"PROVIENE DE PROYECTO", 11),
                        (u"TIPO PROYECTO", 11),
                        (u"TÍTULO DEL PROYECTO", 39),
                        (u"PERTENECE GRUPO INVESTIGACIÓN", 11),
                        (u"GRUPO DE INVESTIGACIÓN", 39),
                        (u"ESTADO", 11),
                        (u"CÉDULA", 11),
                        (u"TIPO PARTICIPANTE", 11),
                        (u"PARTICIPANTE", 58),
                        (u"TIPO PARTICIPACION", 15),
                        (u"TIPO UNEMI", 15),
                        (u"DOI", 23),
                        (u"CUARTIL", 12),
                        (u"JCR", 12),
                        (u"SJR", 12),
                        (u"CATEGORÍA", 23),
                        (u"ENLACE ARTÍCULO", 58),
                        (u"ENLACE REVISTA", 58)
                    ]

                    row_num = 5
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], fcabeceracolumna)
                        ws.set_column(col_num, col_num, columns[col_num][1])

                    row_num = 6

                    if tipo == 'gen':
                        listaarticulos = ParticipantesArticulos.objects.filter(status=True).order_by('articulo__nombre')
                    else:
                        lista_ids_personas = participantes.split(",")
                        listaarticulos = ParticipantesArticulos.objects.filter(Q(profesor__persona__id__in=lista_ids_personas) |
                                                                               Q(administrativo__persona__id__in=lista_ids_personas) |
                                                                               Q(inscripcion__persona__id__in=lista_ids_personas), status=True).order_by('articulo__nombre')

                    for articulos in listaarticulos:
                        listabasesindexadas = ','.join([x.baseindexada.nombre for x in ArticulosBaseIndexada.objects.filter(articulo=articulos.articulo.id, status=True)])

                        if articulos.profesor:
                            nombres = articulos.profesor.persona.nombre_completo_inverso()
                            tipoparticipante = articulos.get_tipoparticipanteins_display()
                            cedula = articulos.profesor.persona.cedula
                        elif articulos.administrativo:
                            nombres = articulos.administrativo.persona.nombre_completo_inverso()
                            tipoparticipante = articulos.get_tipoparticipanteins_display()
                            cedula = articulos.administrativo.persona.cedula
                        elif articulos.inscripcion:
                            nombres = articulos.inscripcion.persona.nombre_completo_inverso()
                            tipoparticipante = articulos.get_tipoparticipanteins_display()
                            cedula = articulos.inscripcion.persona.cedula

                        ws.write(row_num, 0, "REVISTA", fceldageneral)
                        ws.write(row_num, 1, articulos.articulo.revista.codigoissn + ' ' + str(articulos.articulo.id) + '-ART', fceldageneral)
                        ws.write(row_num, 2, articulos.articulo.nombre.strip(), fceldageneral)
                        ws.write(row_num, 3, articulos.articulo.volumen, fceldageneral)
                        ws.write(row_num, 4, articulos.articulo.numero, fceldageneral)
                        ws.write(row_num, 5, articulos.articulo.paginas, fceldageneral)
                        ws.write(row_num, 6, articulos.articulo.revista.codigoissn, fceldageneral)
                        ws.write(row_num, 7, "SI" if articulos.articulo.indexada else "NO", fceldageneralcent)
                        ws.write(row_num, 8, listabasesindexadas, fceldageneral)
                        ws.write(row_num, 9, articulos.articulo.revista.nombre.strip(), fceldageneral)
                        ws.write(row_num, 10, articulos.articulo.revista.get_tipo_display(), fceldageneralcent)
                        ws.write(row_num, 11, articulos.articulo.revista.sjr, fceldageneral)
                        ws.write(row_num, 12, articulos.articulo.fechapublicacion, fceldafechaDMA)
                        ws.write(row_num, 13, articulos.articulo.areaconocimiento.nombre, fceldageneral)
                        ws.write(row_num, 14, articulos.articulo.subareaconocimiento.nombre, fceldageneral)
                        ws.write(row_num, 15, articulos.articulo.subareaespecificaconocimiento.nombre, fceldageneral)
                        ws.write(row_num, 16, articulos.articulo.lineainvestigacion.nombre if articulos.articulo.lineainvestigacion else '', fceldageneral)
                        ws.write(row_num, 17, articulos.articulo.sublineainvestigacion.nombre if articulos.articulo.sublineainvestigacion else '', fceldageneral)
                        ws.write(row_num, 18, 'SI' if articulos.articulo.tipoproyecto else 'NO', fceldageneralcent)
                        ws.write(row_num, 19, articulos.articulo.get_tipoproyecto_display() if articulos.articulo.tipoproyecto else '', fceldageneralcent)

                        if articulos.articulo.tipoproyecto:
                            ws.write(row_num, 20, articulos.articulo.proyectointerno.nombre if articulos.articulo.proyectointerno else articulos.articulo.proyectoexterno.nombre, fceldageneral)
                        else:
                            ws.write(row_num, 20, '', fceldageneral)

                        ws.write(row_num, 21, 'SI' if articulos.articulo.pertenecegrupoinv else 'NO', fceldageneral)
                        ws.write(row_num, 22, articulos.articulo.grupoinvestigacion.nombre if articulos.articulo.pertenecegrupoinv else '', fceldageneral)
                        ws.write(row_num, 23, articulos.articulo.get_estado_display(), fceldageneral)
                        ws.write(row_num, 24, cedula, fceldageneral)
                        ws.write(row_num, 25, articulos.get_tipo_display(), fceldageneral)
                        ws.write(row_num, 26, nombres, fceldageneral)
                        ws.write(row_num, 27, tipoparticipante, fceldageneral)
                        ws.write(row_num, 28, articulos.tipo_unemi(), fceldageneral)
                        ws.write(row_num, 29, articulos.articulo.doy, fceldageneral)
                        ws.write(row_num, 30, articulos.articulo.get_cuartil_display(), fceldageneral)
                        ws.write(row_num, 31, articulos.articulo.jcr, fceldageneral)
                        ws.write(row_num, 32, articulos.articulo.sjr, fceldageneral)
                        ws.write(row_num, 33, articulos.articulo.get_categoria_display(), fceldageneral)
                        ws.write(row_num, 34, articulos.articulo.enlace, fceldageneral)
                        ws.write(row_num, 35, articulos.articulo.revista.enlace, fceldageneral)

                        row_num += 1


                    workbook.close()

                    ruta = "media/postgrado/" + nombrearchivo
                    return JsonResponse({'result': 'ok', 'archivo': ruta})
                except Exception as ex:
                    msg = ex.__str__()
                    return JsonResponse({"result": "bad", "titulo": "Error", "mensaje": u"Error al generar el reporte. [%s]" % msg, "showSwal": "True", "swalType": "error"})

            elif action == 'addgrupo':
                try:
                    data['title'] = u'Agregar Grupo de Investigación'
                    form = GrupoInvestigacionForm()
                    data['form'] = form
                    return render(request, "ges_grupoinvestigacion/addgrupo.html", data)
                except Exception as ex:
                    pass

            elif action == 'addintegrante':
                try:
                    data['title'] = u'Agregar Integrante al Grupo de Investigación'
                    data['funciones'] = FUNCION_INTEGRANTE_GRUPO_INVESTIGACION

                    template = get_template("ges_grupoinvestigacion/addintegrante.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content, 'title': data['title']})
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

            elif action == 'editgrupo':
                try:
                    data['title'] = u'Editar Grupo de Investigación'
                    grupoinvestigacion = GrupoInvestigacion.objects.get(pk=int(encrypt(request.GET['id'])))

                    form = GrupoInvestigacionForm(
                        initial={
                            'nombre': grupoinvestigacion.nombre,
                            'descripcion': grupoinvestigacion.descripcion,
                            'vigente': grupoinvestigacion.vigente
                        }
                    )

                    data['form'] = form
                    data['id'] = request.GET['id']
                    data['integrantes'] = integrantes = grupoinvestigacion.integrantes()
                    data['totalintegrantes'] = integrantes.count()
                    data['funciones'] = FUNCION_INTEGRANTE_GRUPO_INVESTIGACION

                    return render(request, "ges_grupoinvestigacion/editgrupo.html", data)
                except Exception as ex:
                    pass

            elif action == 'solicitudesgrupo':
                try:
                    search, url_vars, ids, idi, tinf = request.GET.get('s', ''), '&action=' + action, request.GET.get('ids', ''), request.GET.get('idi', ''), request.GET.get('tinf', '')

                    if persona.es_vicerrector_investigacion():
                        filtro = Q(status=True, aprobadocoord=True, grupoinvestigacionrecorrido__estado__valor=5)
                    elif persona.es_coordinador_investigacion():
                        filtro = Q(status=True, aprobadocoord=True, grupoinvestigacionrecorrido__estado__valor=7)
                    elif persona.es_tecnico_investigacion() or persona.es_experto_investigacion():
                        filtro = Q(status=True, aprobadocoord=True, grupoinvestigacionrecorrido__estado__valor=8)
                    else:
                        filtro = Q(status=True, aprobadocoord=True, aprobadoocs=True, grupoinvestigacionrecorrido__estado__valor=1)

                    if search:
                        data['s'] = search
                        filtro = filtro & (Q(nombre__unaccent__icontains=search))
                        url_vars += '&s=' + search

                    if ids:
                        data['ids'] = ids
                        filtro = filtro & Q(pk=encrypt(ids))
                        url_vars += '&ids=' + ids

                    if idi:
                        informe = GrupoInvestigacionInforme.objects.get(pk=int(encrypt(idi)))
                        data['informe'] = informe.archivo.url if tinf == 'sf' else informe.archivofirmado.url
                        data['tipoinforme'] = 'Informe Técnico de Grupo de Investigación' if tinf == 'sf' else 'Informe Técnico de Grupo de Investigación Firmado'

                    gruposinvestigacion = GrupoInvestigacion.objects.filter(filtro).distinct().order_by('-numero')

                    # Determinar si existen informes aprobados y por remitir a CGA
                    informeaprobado = GrupoInvestigacionInforme.objects.values("id").filter(status=True, firmaaprueba=True).exists() if persona.es_vicerrector_investigacion() or persona.es_tecnico_investigacion() else False
                    informexremitir = GrupoInvestigacionInforme.objects.values("id").filter(status=True, firmaaprueba=True, remitidocga=False).exists() if persona.es_vicerrector_investigacion() or persona.es_tecnico_investigacion() else False

                    paging = MiPaginador(gruposinvestigacion, 25)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['url_vars'] = url_vars
                    data['gruposinvestigacion'] = page.object_list
                    data['informeaprobado'] = informeaprobado
                    data['informexremitir'] = informexremitir
                    data['title'] = u'Solicitudes de Propuestas de Creación de Grupos de Investigación'
                    data['enlaceatras'] = "/ges_grupoinvestigacion"

                    return render(request, "ges_grupoinvestigacion/solicitud.html", data)
                except Exception as ex:
                    pass

            elif action == 'mostrarrecorrido':
                try:
                    title = u'Recorrido de la Solicitud de Propuesta de Creación de Grupo de Investigación'
                    grupoinvestigacion = GrupoInvestigacion.objects.get(pk=int(encrypt(request.GET['id'])))
                    data['grupoinvestigacion'] = grupoinvestigacion
                    data['recorrido'] = grupoinvestigacion.recorrido()
                    template = get_template("pro_grupoinvestigacion/recorrido.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content, 'title': title})
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

            elif action == 'informacion':
                try:
                    data['title'] = u'Información Solicitud de Propuesta de Creación de Grupo de Investigación'
                    data['grupo'] = grupoinvestigacion = GrupoInvestigacion.objects.get(pk=int(encrypt(request.GET['id'])))
                    data['objetivos'] = grupoinvestigacion.objetivos_especificos()
                    data['lineasinvestigacion'] = grupoinvestigacion.lineas_grupo()
                    data['integrantes'] = grupoinvestigacion.integrantes()
                    data['tecnologias'] = grupoinvestigacion.tecnologias()
                    data['requisitos'] = grupoinvestigacion.requisitos_director()
                    data['resolucion'] = grupoinvestigacion.resolucion_facultad()
                    data['resolucionocs'] = grupoinvestigacion.resolucion_ocs()
                    data['modulo'] = request.GET['mod']
                    return render(request, "pro_grupoinvestigacion/informacion.html", data)
                except Exception as ex:
                    pass

            elif action == 'reasignarsolicitud':
                try:
                    data['title'] = u'Reasignar Solicitud para Análisis (Al Coordinador de Investigación)' if request.GET['tipodest'] == 'ci' else u'Reasignar Solicitud para Análisis (Al Analista de Investigación)'
                    data['subtitle'] = u'Solicitud de Propuesta para Creación de Grupo de Investigación'
                    data['grupo'] = grupoinvestigacion = GrupoInvestigacion.objects.get(pk=int(encrypt(request.GET['id'])))
                    data['objetivos'] = grupoinvestigacion.objetivos_especificos()
                    data['lineasinvestigacion'] = grupoinvestigacion.lineas_grupo()
                    data['integrantes'] = grupoinvestigacion.integrantes()
                    data['tecnologias'] = grupoinvestigacion.tecnologias()
                    data['requisitos'] = grupoinvestigacion.requisitos_director()
                    data['destinatario'] = coordinador_investigacion() if request.GET['tipodest'] == 'ci' else tecnico_revisor_grupoinvestigacion()
                    data['tipodest'] = request.GET['tipodest']
                    data['resolucion'] = grupoinvestigacion.resolucion_facultad()
                    return render(request, "ges_grupoinvestigacion/reasignarsolicitud.html", data)
                except Exception as ex:
                    pass

            elif action == 'analizarsolicitud':
                try:
                    data['title'] = u'Analizar Solicitud de Propuesta de Creación de Grupo de Investigación'
                    data['grupo'] = grupoinvestigacion = GrupoInvestigacion.objects.get(pk=int(encrypt(request.GET['id'])))
                    data['objetivos'] = grupoinvestigacion.objetivos_especificos()
                    data['lineasinvestigacion'] = grupoinvestigacion.lineas_grupo()
                    data['integrantes'] = grupoinvestigacion.integrantes()
                    data['tecnologias'] = grupoinvestigacion.tecnologias()
                    data['requisitos'] = grupoinvestigacion.requisitos_director()
                    data['resolucion'] = grupoinvestigacion.resolucion_facultad()
                    data['estados'] = obtener_estados_solicitud(19, [9, 10])
                    return render(request, "ges_grupoinvestigacion/analizarsolicitudgrupo.html", data)
                except Exception as ex:
                    pass

            elif action == 'addinforme':
                try:
                    data['title'] = u'Agregar Informe Técnico de Creación de Grupo de Investigación'
                    data['grupo'] = grupoinvestigacion = GrupoInvestigacion.objects.get(pk=int(encrypt(request.GET['id'])))
                    data['objetivos'] = grupoinvestigacion.objetivos_especificos()
                    data['lineasinvestigacion'] = grupoinvestigacion.lineas_grupo()
                    data['integrantes'] = grupoinvestigacion.integrantes()
                    data['tecnologias'] = grupoinvestigacion.tecnologias()
                    data['requisitos'] = grupoinvestigacion.requisitos_director()
                    data['resolucion'] = grupoinvestigacion.resolucion_facultad()
                    data['coordinador'] = coordinador_investigacion()
                    data['director'] = vicerrector_investigacion_posgrado()
                    data['fecha'] = datetime.now().date()

                    return render(request, "ges_grupoinvestigacion/addinforme.html", data)
                except Exception as ex:
                    pass

            elif action == 'editinforme':
                try:
                    data['title'] = u'Editar Informe Técnico de Creación de Grupo de Investigación'
                    data['informe'] = informe = GrupoInvestigacionInforme.objects.get(pk=int(encrypt(request.GET['id'])))
                    data['grupo'] = grupoinvestigacion = informe.grupo
                    data['objetivos'] = grupoinvestigacion.objetivos_especificos()
                    data['lineasinvestigacion'] = grupoinvestigacion.lineas_grupo()
                    data['integrantes'] = grupoinvestigacion.integrantes()
                    data['tecnologias'] = grupoinvestigacion.tecnologias()
                    data['requisitos'] = grupoinvestigacion.requisitos_director()
                    data['resolucion'] = grupoinvestigacion.resolucion_facultad()
                    data['coordinador'] = coordinador_investigacion()
                    data['director'] = vicerrector_investigacion_posgrado()
                    data['anexos'] = anexos =informe.anexos()
                    data['totalanexos'] = anexos.count()
                    data['fecha'] = datetime.now().date()

                    return render(request, "ges_grupoinvestigacion/editinforme.html", data)
                except Exception as ex:
                    pass

            elif action == 'firmarinforme':
                try:
                    tipofirma = request.GET['tipofirma']
                    data['informe'] = informe = GrupoInvestigacionInforme.objects.get(pk=int(encrypt(request.GET['id'])))
                    data['iddoc'] = informe.id
                    data['tipofirma'] = tipofirma

                    grupoinvestigacion = informe.grupo

                    if tipofirma == 'ELA':  # Persona que elabora
                        data['title'] = u'Firmar Informe Técnico. Elaborado por: {}'.format(informe.elabora.nombre_completo_inverso())
                        data['idper'] = informe.elabora.id
                    elif tipofirma == 'VER':  # Persona que verifica
                        data['title'] = u'Firmar Informe Técnico. Verificado por: {}'.format(informe.verifica.nombre_completo_inverso())
                        data['idper'] = informe.verifica.id
                    else:  # Persona que aprueba
                        data['title'] = u'Firmar Informe Técnico. Aprobado por: {}'.format(informe.aprueba.nombre_completo_inverso())
                        data['idper'] = informe.aprueba.id

                    data['mensaje'] = "Firma de Informe Técnico de Creación de Grupo de Investigación N° <b>{}</b> del docente <b>{}</b> para el grupo de investigación <b>{}</b>".format(informe.numero, grupoinvestigacion.profesor.persona.nombre_completo_inverso(), grupoinvestigacion.nombre)
                    data['accionfirma'] = action

                    template = get_template("pro_becadocente/firmardocumento.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content, 'title': data['title']})
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

            elif action == 'validarinforme':
                try:
                    data['title'] = u'Analizar Informe Técnico de Creación de Grupo de Investigación'
                    data['informe'] = GrupoInvestigacionInforme.objects.get(pk=int(encrypt(request.GET['id'])))

                    estados = []
                    estados.append({"id": 2, "descripcion": "VALIDADO"})
                    estados.append({"id": 3, "descripcion": "NOVEDAD"})
                    data['estados'] = estados

                    template = get_template("ges_grupoinvestigacion/validarinforme.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content, 'title': data['title']})
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

            elif action == 'evidenciasarticulo':
                try:
                    data['aniosevidencias'] = ArticuloInvestigacion.objects.filter(status=True, fechapublicacion__isnull=False).annotate(anio=ExtractYear('fechapublicacion')).values_list('anio', flat=True).order_by('-anio').distinct()
                    data['action'] = 'evidenciasarticulo'
                    template = get_template("pcientifica/modal/descargarevidencia.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

            elif action == 'evidenciasarticulocodigo':
                try:
                    form = EvidenciasCodigoForm()
                    data['form'] = form
                    template = get_template("pcientifica/modal/descargarevidenciacodigo.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    msg = ex.__str__()
                    return JsonResponse({"result": False, "titulo": "Error", "mensaje": f'Error al obtener los datos. {msg}'}, safe=False)

            elif action == 'integrantes':
                try:
                    title = u'Seleccionar Participantes'
                    data['participantes'] = Persona.objects.values('id', 'cedula', 'nombres', 'apellido1', 'apellido2').filter(Q(profesor__participantesarticulos__isnull=False, profesor__participantesarticulos__status=True))
                    template = get_template("pcientifica/modal/seleccionparticipantes.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content, 'title': title})
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

            elif action == 'solicitudarticulo':
                try:
                    search, filtro, url_vars = request.GET.get('s', ''), Q(tiposolicitud__in=[1, 5], status=True, estado__valor=1), f'&action={action}'

                    if search:
                        data['s'] = search
                        filtro = filtro & (Q(nombre__unaccent__icontains=search))
                        url_vars += '&s=' + search

                    solicitudes = SolicitudPublicacion.objects.filter(filtro).order_by('fecha_creacion')

                    paging = MiPaginador(solicitudes, 25)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['url_vars'] = url_vars
                    data['solicitudes'] = page.object_list
                    data['title'] = u'Solicitudes de Registro de Artículos'

                    request.session['viewactivopc'] = ['mnuarticulos', 'solicitudesart']
                    return render(request, "pcientifica/viewsolicitudarticulo.html", data)
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'revistas':
                try:
                    data['title'] = u'Revistas'
                    data['persona_'] = persona = Persona.objects.get(id=persona.id)
                    data['perfil'] = persona.mi_perfil()
                    # data['migrante'] = MigrantePersona.objects.filter(persona=persona).first()
                    data['documentopersonal'] = persona.documentos_personales()
                    # data['reporte_1'] = obtener_reporte('hoja_vida_sagest')
                    # data['reporte_2'] = obtener_reporte('certificado_ficha_estudiantil')

                    request.session['viewactivopc'] = ['mnuarticulos', 'revistas']
                    return render(request, "pcientifica/viewrevista.html", data)
                except Exception as ex:
                    # msg = ex.__str__()
                    messages.error(request, f'{ex}')
                    # return JsonResponse({"result": "bad", "titulo": "Error", "mensaje": u"Error al obtener los datos. [%s]" % msg, "showSwal": "True", "swalType": "error"})

            elif action == 'bases':
                try:
                    data['title'] = u'Bases Indexadas'
                    data['persona_'] = persona = Persona.objects.get(id=persona.id)
                    data['perfil'] = persona.mi_perfil()
                    # data['migrante'] = MigrantePersona.objects.filter(persona=persona).first()
                    data['documentopersonal'] = persona.documentos_personales()
                    # data['reporte_1'] = obtener_reporte('hoja_vida_sagest')
                    # data['reporte_2'] = obtener_reporte('certificado_ficha_estudiantil')

                    request.session['viewactivopc'] = ['mnuarticulos', 'bases']
                    return render(request, "pcientifica/viewbase.html", data)
                except Exception as ex:
                    # msg = ex.__str__()
                    messages.error(request, f'{ex}')
                    # return JsonResponse({"result": "bad", "titulo": "Error", "mensaje": u"Error al obtener los datos. [%s]" % msg, "showSwal": "True", "swalType": "error"})

            elif action == 'perfilesorcid':
                try:
                    data['title'] = u'Perfiles ORCID'
                    data['persona_'] = persona = Persona.objects.get(id=persona.id)
                    data['perfil'] = persona.mi_perfil()
                    # data['migrante'] = MigrantePersona.objects.filter(persona=persona).first()
                    data['documentopersonal'] = persona.documentos_personales()
                    # data['reporte_1'] = obtener_reporte('hoja_vida_sagest')
                    # data['reporte_2'] = obtener_reporte('certificado_ficha_estudiantil')

                    request.session['viewactivopc'] = ['mnuarticulos', 'perfilesorcid']
                    return render(request, "pcientifica/viewperfilorcid.html", data)
                except Exception as ex:
                    # msg = ex.__str__()
                    messages.error(request, f'{ex}')
                    # return JsonResponse({"result": "bad", "titulo": "Error", "mensaje": u"Error al obtener los datos. [%s]" % msg, "showSwal": "True", "swalType": "error"})

            elif action == 'perfilesscopus':
                try:
                    data['title'] = u'Perfiles SCOPUS'
                    data['persona_'] = persona = Persona.objects.get(id=persona.id)
                    data['perfil'] = persona.mi_perfil()
                    # data['migrante'] = MigrantePersona.objects.filter(persona=persona).first()
                    data['documentopersonal'] = persona.documentos_personales()
                    # data['reporte_1'] = obtener_reporte('hoja_vida_sagest')
                    # data['reporte_2'] = obtener_reporte('certificado_ficha_estudiantil')

                    request.session['viewactivopc'] = ['mnuarticulos', 'perfilesscopus']
                    return render(request, "pcientifica/viewperfilscopus.html", data)
                except Exception as ex:
                    # msg = ex.__str__()
                    messages.error(request, f'{ex}')
                    # return JsonResponse({"result": "bad", "titulo": "Error", "mensaje": u"Error al obtener los datos. [%s]" % msg, "showSwal": "True", "swalType": "error"})

            elif action == 'ponencias':
                try:
                    data['title'] = u'Ponencias'
                    data['persona_'] = persona = Persona.objects.get(id=persona.id)
                    data['perfil'] = persona.mi_perfil()
                    # data['migrante'] = MigrantePersona.objects.filter(persona=persona).first()
                    data['documentopersonal'] = persona.documentos_personales()
                    # data['reporte_1'] = obtener_reporte('hoja_vida_sagest')
                    # data['reporte_2'] = obtener_reporte('certificado_ficha_estudiantil')

                    request.session['viewactivopc'] = ['mnuponencias', 'ponencias']
                    return render(request, "pcientifica/viewponencia.html", data)
                except Exception as ex:
                    # msg = ex.__str__()
                    messages.error(request, f'{ex}')
                    # return JsonResponse({"result": "bad", "titulo": "Error", "mensaje": u"Error al obtener los datos. [%s]" % msg, "showSwal": "True", "swalType": "error"})

            elif action == 'solicitudponencia':
                try:
                    data['title'] = u'Solicitudes de Ponencias'
                    data['persona_'] = persona = Persona.objects.get(id=persona.id)
                    data['perfil'] = persona.mi_perfil()
                    # data['migrante'] = MigrantePersona.objects.filter(persona=persona).first()
                    data['documentopersonal'] = persona.documentos_personales()
                    # data['reporte_1'] = obtener_reporte('hoja_vida_sagest')
                    # data['reporte_2'] = obtener_reporte('certificado_ficha_estudiantil')

                    request.session['viewactivopc'] = ['mnuponencias', 'solicitudponencia']
                    return render(request, "pcientifica/viewsolicitudponencia.html", data)
                except Exception as ex:
                    # msg = ex.__str__()
                    messages.error(request, f'{ex}')
                    # return JsonResponse({"result": "bad", "titulo": "Error", "mensaje": u"Error al obtener los datos. [%s]" % msg, "showSwal": "True", "swalType": "error"})


            return HttpResponseRedirect(request.path)
        else:
            try:
                search, id, estado, buscarpor, idsp, filtro, url_vars = request.GET.get('s', ''), request.GET.get('id', ''), request.GET.get('estado', ''), request.GET.get('buscarpor', ''), request.GET.get('idsp', ''), Q(status=True) | Q(status=False, eliminadoxdoc=True), ''

                if id:
                    filtro = filtro & (Q(pk=int(encrypt(id))))
                    url_vars += '&id=' + id
                else:
                    if search:
                        idsp = ''
                        data['s'] = search
                        data['buscarpor'] = buscarpor = int(buscarpor)

                        if buscarpor == 1:
                            filtro = filtro & (Q(nombre__unaccent__icontains=search))
                        elif buscarpor == 2:
                            filtro = filtro & (Q(revista__nombre__unaccent__icontains=search))
                        elif buscarpor == 3:
                            if ' ' not in search:
                                articulosparticipantesid = ParticipantesArticulos.objects.values_list('articulo_id').filter(status=True).filter(
                                    Q(profesor__persona__nombres__icontains=search) |
                                    Q(profesor__persona__apellido1__icontains=search) |
                                    Q(profesor__persona__apellido2__icontains=search) |
                                    Q(administrativo__persona__nombres__icontains=search) |
                                    Q(administrativo__persona__apellido1__icontains=search) |
                                    Q(administrativo__persona__apellido2__icontains=search) |
                                    Q(inscripcion__persona__nombres__icontains=search) |
                                    Q(inscripcion__persona__apellido1__icontains=search) |
                                    Q(inscripcion__persona__apellido2__icontains=search)).distinct()
                            else:
                                s = search.split(" ")
                                articulosparticipantesid = ParticipantesArticulos.objects.values_list('articulo_id').filter(status=True).filter(
                                    (Q(profesor__persona__apellido1__icontains=s[0]) &
                                    Q(profesor__persona__apellido2__icontains=s[1])) |
                                    (Q(administrativo__persona__apellido1__icontains=s[0]) &
                                    Q(administrativo__persona__apellido2__icontains=s[1]))|
                                    (Q(inscripcion__persona__apellido1__icontains=s[0]) &
                                    Q(inscripcion__persona__apellido2__icontains=s[1]))).distinct()

                            filtro = filtro & (Q(pk__in=articulosparticipantesid))
                        elif buscarpor == 4:
                            filtro = filtro & (Q(articulosbaseindexada__baseindexada__nombre__unaccent__icontains=search))
                        elif buscarpor == 5:
                            filtro = filtro & (Q(fechapublicacion__year=search))

                        url_vars += '&s=' + search + '&buscarpor=' + str(buscarpor)

                    if estado:
                        data['estado'] = estado
                        filtro = filtro & (Q(aprobado=True) if int(estado) == 1 else Q(aprobado=False))
                        url_vars += '&estado=' + estado

                    if idsp:
                        data['buscarpor'] = buscarpor = int(buscarpor)
                        lista_ids_personas = idsp.split(",")
                        articulosparticipantesid = ParticipantesArticulos.objects.values_list('articulo_id').filter(status=True).filter(Q(profesor__persona__id__in=lista_ids_personas) |
                                                                                                                                        Q(administrativo__persona__id__in=lista_ids_personas) |
                                                                                                                                        Q(inscripcion__persona__id__in=lista_ids_personas)).distinct()

                        filtro = filtro & (Q(pk__in=articulosparticipantesid))
                        url_vars += '&idsp=' + idsp + '&buscarpor=' + str(buscarpor)

                articulos = ArticuloInvestigacion.objects.select_related().filter(filtro).order_by('-fechapublicacion', 'revista__nombre', 'numero', 'nombre')

                totalarticulos = articulos.count()
                totalaprobados = articulos.filter(aprobado=True).count()
                totalporaprobar = totalarticulos - totalaprobados

                tiposbusqueda = [
                    {"id": 1, "descripcion": "Artículo"},
                    {"id": 2, "descripcion": "Revista"},
                    {"id": 3, "descripcion": "Participantes"},
                    {"id": 4, "descripcion": "Base Indexada"},
                    {"id": 5, "descripcion": "Año Publicación"},
                ]

                paging = MiPaginador(articulos, 25)
                p = 1
                try:
                    paginasesion = 1
                    if 'paginador' in request.session:
                        paginasesion = int(request.session['paginador'])
                    if 'page' in request.GET:
                        p = int(request.GET['page'])
                    else:
                        p = paginasesion
                    try:
                        page = paging.page(p)
                    except:
                        p = 1
                    page = paging.page(p)
                except:
                    page = paging.page(p)
                request.session['paginador'] = p
                data['paging'] = paging
                data['rangospaging'] = paging.rangos_paginado(p)
                data['page'] = page
                data['url_vars'] = url_vars
                data['articulos'] = page.object_list
                data['totalarticulos'] = totalarticulos
                data['totalaprobados'] = totalaprobados
                data['totalporaprobar'] = totalporaprobar
                data['tiposbusqueda'] = tiposbusqueda
                data['title'] = u'Artículos'

                request.session['viewactivopc'] = ['mnuarticulos', 'articulos']
                return render(request, "pcientifica/viewarticulo.html", data)
            except Exception as ex:
                pass
