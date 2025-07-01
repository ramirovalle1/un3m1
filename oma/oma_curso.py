import io
import json
import os
import sys
import pyqrcode
from decimal import Decimal
from datetime import datetime, timedelta
import openpyxl
import xlsxwriter
import xlwt
from django.core.exceptions import ObjectDoesNotExist
from xlwt import *
import random
import shutil
from django.db.models import Sum
from django.db.models import Count
from googletrans import Translator
from django.contrib.auth.decorators import login_required
from django.db import transaction, connection
from django.db.models import Q, Max
from django.forms import model_to_dict
from django.http import HttpResponseRedirect, JsonResponse, HttpResponse
from django.shortcuts import render, redirect
from django.template import Context
from django.template.loader import get_template
from openpyxl import load_workbook
from decorators import secure_module
from settings import RESPONSABLE_BIENES_ID, MEDIA_ROOT, ASISTENTE_BODEGA_ID, SITE_ROOT, JR_USEROUTPUT_FOLDER, \
    SITE_STORAGE, JR_JAVA_COMMAND, \
    JR_RUN, DATABASES, SUBREPOTRS_FOLDER, MEDIA_URL, DEBUG
from api.helpers.functions_helper import get_variable, remove_accents, transform_jasperstarter, fixparametro, \
    fetch_resources
from sga.commonviews import adduserdata, obtener_reporte
from sga.forms import ReservasCraiSolicitarAutoridadForm
from sga.templatetags.sga_extras import encrypt
from sga.funciones import MiPaginador, generar_nombre, log, convertir_fecha, convertir_fecha_invertida, \
    puede_realizar_accion, remover_caracteres_especiales_unicode, null_to_decimal, remover_caracteres_tildes_unicode, \
    convertir_lista, null_to_numeric
from sga.models import Persona, Reporte
from sga.reportes import elimina_tildes, run_report_v1
from sga.funcionesxhtml2pdf import conviert_html_to_pdf, conviert_html_to_pdf_name, \
    conviert_html_to_pdfsaveqr_omacertificado
from core.firmar_documentos import firmar, firmararchivogenerado
from oma.models import Curso, InscripcionCurso, AsignaturaCurso, AsignaturaInscripcionCurso, EvaluacionGenerica, \
    DetalleModeloEvaluativo, AuditoriaNotasOma, ModeloEvaluativo
from oma.forms import CursoForm, InscripcionCursoForm, AsignaturaCursoForm, MasivoInscripcionForm, ModeloEvaluativoForm, \
    LogicaModeloEvaluativoForm, DetalleModeloEvaluativoForm
from sga.models import Inscripcion, Asignatura, Matricula, Periodo, CUENTAS_CORREOS, RecordAcademico,\
    HistoricoRecordAcademico
from sga.funciones import variable_valor
from sga.tasks import send_html_mail

unicode = str


@login_required(redirect_field_name='ret', login_url='/loginsga')
# @secure_module
# @transaction.atomic()
def view(request):
    data = {}
    adduserdata(request, data)
    usuario = request.user
    persona = request.session['persona']

    h = 'https'
    if DEBUG:
        h = 'http'
    base_url = request.META['HTTP_HOST']
    data['DOMINIO_DEL_SISTEMA'] = dominio_sistema = f"{h}://{unicode(base_url)}"

    data['IS_DEBUG'] = IS_DEBUG = variable_valor('IS_DEBUG')

    if request.method == 'POST':
        action = request.POST['action']

        if action == 'add':
            try:
                f = CursoForm(request.POST)
                if f.is_valid():
                    if Curso.objects.filter(nombre=f.cleaned_data['nombre'].upper().strip(),
                                            fecha_inicio=f.cleaned_data['fechainicio'],
                                            fecha_fin=f.cleaned_data['fechafin'],
                                            status=True).exists():
                        return JsonResponse({"result": True, "mensaje": u"El registro ya existe."})
                    curso = Curso(nombre=f.cleaned_data['nombre'],
                                  codigo=f.cleaned_data['codigo'],
                                  modeloevaluativo=f.cleaned_data['modeloevaluativo'],
                                  fecha_inicio=f.cleaned_data['fechainicio'],
                                  fecha_fin=f.cleaned_data['fechafin'],
                                  horas=f.cleaned_data['horas'],
                                  creditos=f.cleaned_data['creditos'],
                                  enlacegrabacion=f.cleaned_data['enlacegrabacion'],
                                  enlacereuniondocente=f.cleaned_data['enlacereuniondocente'],
                                  idcursomoodle=f.cleaned_data['idcursomoodle'],
                                  enlacepresentacioncurso=f.cleaned_data['enlacepresentacioncurso'])
                    curso.save(request)

                    log(u'Adicionó nuevo curso - oma: %s' % curso, request, "add")
                    return JsonResponse({"result": False, 'mensaje': 'Registro Exitoso'})
                else:
                    # print([{k: v[0]} for k, v in f.errors.items()])
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": u"Error al guardar los datos."})

        elif action == 'addinscripcioncurso':
            try:
                f = InscripcionCursoForm(request.POST)
                if f.is_valid():
                    curso = Curso.objects.get(pk=int(encrypt(request.POST['id'])))
                    inscripcion = f.cleaned_data['inscripcion']
                    if InscripcionCurso.objects.filter(curso=curso, inscripcion=inscripcion, status=True).exists():
                        return JsonResponse({"result": True, "mensaje": u"El registro ya existe."})
                    if InscripcionCurso.objects.filter(curso=curso, inscripcion__persona=inscripcion.persona,
                                                       status=True).exists():
                        return JsonResponse({"result": True, "mensaje": u"Ya existe registro con esta persona."})
                    insccurso = InscripcionCurso(curso_id=int(encrypt(request.POST['id'])),
                                                 inscripcion=f.cleaned_data['inscripcion'],
                                                 correo=f.cleaned_data['correo'])
                    insccurso.save(request)

                    for x in AsignaturaCurso.objects.filter(curso=curso, status=True):
                        try:
                            asiginsccurso = AsignaturaInscripcionCurso.objects.get(asignaturacurso=x,
                                                                                   inscripcioncurso=insccurso,
                                                                                   status=True)
                        except ObjectDoesNotExist:
                            asiginsccurso = AsignaturaInscripcionCurso(asignaturacurso=x, inscripcioncurso=insccurso)
                            asiginsccurso.save(request)

                        for y in DetalleModeloEvaluativo.objects.filter(modelo=curso.modeloevaluativo,
                                                                        status=True):
                            try:
                                evaluagenerica = EvaluacionGenerica.objects.get(detallemodeloevaluativo=y,
                                                                                asignaturainscripcion=asiginsccurso,
                                                                                status=True)
                            except ObjectDoesNotExist:
                                evaluagenerica = EvaluacionGenerica(detallemodeloevaluativo=y,
                                                                    asignaturainscripcion=asiginsccurso)
                                evaluagenerica.save(request)

                    log(u'Adicionó nueva inscripcion curso - oma: %s' % curso, request, "add")
                    return JsonResponse({"result": False, 'mensaje': 'Registro Exitoso'})
                else:
                    # print([{k: v[0]} for k, v in f.errors.items()])
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": u"Error al guardar los datos."})

        elif action == 'addasignaturacurso':
            try:
                f = AsignaturaCursoForm(request.POST)
                if f.is_valid():
                    if AsignaturaCurso.objects.filter(curso_id=int(encrypt(request.POST['id'])),
                                                      asignatura=f.cleaned_data['asignatura'], status=True).exists():
                        return JsonResponse({"result": True, "mensaje": u"El registro ya existe."})
                    asig = AsignaturaCurso(curso_id=int(encrypt(request.POST['id'])),
                                           asignatura=f.cleaned_data['asignatura'])
                    asig.save(request)

                    log(u'Adicionó nueva asignatura curso - oma: %s' % asig, request, "add")
                    return JsonResponse({"result": False, 'mensaje': 'Registro Exitoso'})
                else:
                    # print([{k: v[0]} for k, v in f.errors.items()])
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": u"Error al guardar los datos."})

        elif action == 'edit':
            try:
                f = CursoForm(request.POST)
                if f.is_valid():
                    curso = Curso.objects.get(pk=encrypt(request.POST['id']))
                    if Curso.objects.filter(nombre=f.cleaned_data['nombre'].upper().strip(),
                                            fecha_inicio=f.cleaned_data['fechainicio'],
                                            fecha_fin=f.cleaned_data['fechafin'],
                                            status=True).exclude(id=curso.id).exists():
                        return JsonResponse({"result": True, "mensaje": u"El registro ya existe."})
                    if Curso.objects.filter(codigo=f.cleaned_data['codigo']).exclude(id=curso.id).exists():
                        return JsonResponse({"result": True, "mensaje": u"Ya existe registro con este código."})
                    # if request.user.has_perm('sga.puede_titulo_tthh'):
                    curso.nombre = f.cleaned_data['nombre']
                    curso.codigo = f.cleaned_data['codigo']
                    curso.fecha_inicio = f.cleaned_data['fechainicio']
                    curso.fecha_fin = f.cleaned_data['fechafin']
                    curso.idcursomoodle = f.cleaned_data['idcursomoodle']
                    # curso.min_aprueba_nota = f.cleaned_data['minnota']
                    curso.horas = f.cleaned_data['horas']
                    curso.creditos = f.cleaned_data['creditos']
                    curso.enlacegrabacion = f.cleaned_data['enlacegrabacion']
                    curso.enlacereuniondocente = f.cleaned_data['enlacereuniondocente']
                    curso.enlacepresentacioncurso = f.cleaned_data['enlacepresentacioncurso']
                    cambiomodelo = False
                    if not curso.modeloevaluativo ==  f.cleaned_data['modeloevaluativo']:
                        cambiomodelo = True
                    curso.modeloevaluativo = f.cleaned_data['modeloevaluativo']
                    curso.save(request)

                    if cambiomodelo:
                        if curso.inscripcioncurso_set.filter(status=True).exists():
                            for inscrito in curso.inscripcioncurso_set.filter(status=True):
                                inscrito.edit_modelo_curso()
                                for x in AsignaturaCurso.objects.filter(curso=curso, status=True):
                                    try:
                                        asiginsccurso = AsignaturaInscripcionCurso.objects.get(asignaturacurso=x,
                                                                                               inscripcioncurso=inscrito,
                                                                                               status=True)
                                    except ObjectDoesNotExist:
                                        asiginsccurso = AsignaturaInscripcionCurso(asignaturacurso=x,
                                                                                   inscripcioncurso=inscrito)
                                        asiginsccurso.save(request)

                                    for y in DetalleModeloEvaluativo.objects.filter(modelo=curso.modeloevaluativo,
                                                                                    status=True):
                                        try:
                                            evaluagenerica = EvaluacionGenerica.objects.get(detallemodeloevaluativo=y,
                                                                                            asignaturainscripcion=asiginsccurso,
                                                                                            status=True)
                                        except ObjectDoesNotExist:
                                            evaluagenerica = EvaluacionGenerica(detallemodeloevaluativo=y,
                                                                                asignaturainscripcion=asiginsccurso)
                                            evaluagenerica.save(request)

                    log(u'Modificó curso ofimática: %s' % curso, request, "edit")
                    return JsonResponse({"result": False, 'mensaje': 'Edicion Exitosa'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": u"Error al guardar los datos."})

        elif action == 'delete':
            try:
                curso = Curso.objects.get(pk=int(encrypt(request.POST['id'])))
                curso.status = False
                curso.save()
                log(u'Elimino curso - oma: %s' % curso, request, "del")
                # return JsonResponse({"result": "ok"})
                res_json = {'error': False}
            except Exception as ex:
                transaction.set_rollback(True)
                res_json = {'error': True, "message": "Error: {}".format(ex)}
                # return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})
            return JsonResponse(res_json, safe=False)

        elif action == 'delinscripcion':
            try:
                inscripcion = InscripcionCurso.objects.get(pk=int(encrypt(request.POST['id'])))
                inscripcion.status = False
                inscripcion.save(request)
                inscripcion.delete_inscripcion()
                log(u'Elimino inscripcion curso - oma: %s' % inscripcion, request, "del")
                # return JsonResponse({"result": "ok"})
                res_json = {'error': False}
            except Exception as ex:
                transaction.set_rollback(True)
                res_json = {'error': True, "message": "Error: {}".format(ex)}
                # return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})
            return JsonResponse(res_json, safe=False)

        elif action == 'delasignatura':
            try:
                asignatura = AsignaturaCurso.objects.get(pk=int(encrypt(request.POST['id'])))
                asignatura.status = False
                asignatura.save(request)
                log(u'Elimino inscripcion curso - oma: %s' % asignatura, request, "del")
                # return JsonResponse({"result": "ok"})
                res_json = {'error': False}
            except Exception as ex:
                transaction.set_rollback(True)
                res_json = {'error': True, "message": "Error: {}".format(ex)}
                # return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})
            return JsonResponse(res_json, safe=False)

        elif action == 'extraernotasmoodle':
            try:
                curso = Curso.objects.get(pk=int(encrypt(request.POST['id'])))
                for alumno in curso.inscripcioncurso_set.filter(status=True):
                    for notasmooc in curso.notas_de_moodle(alumno.inscripcion.persona):
                        for asigInscr in AsignaturaInscripcionCurso.objects.filter(status=True, inscripcioncurso=alumno):
                            campo = asigInscr.campo(notasmooc[1].upper().strip())
                            if type(notasmooc[0]) is Decimal:
                                if null_to_decimal(campo.valor) != float(notasmooc[0]):
                                    actualizar_nota_ofimatica(asigInscr.id, notasmooc[1].upper().strip(), notasmooc[0])
                                    auditorianotas = AuditoriaNotasOma(evaluaciongenerica=campo, manual=False,
                                                                       calificacion=notasmooc[0])
                                    auditorianotas.save()
                            else:
                                if null_to_decimal(campo.valor) != float(0):
                                    actualizar_nota_ofimatica(asigInscr.id, notasmooc[1].upper().strip(), notasmooc[0])
                                    auditorianotas = AuditoriaNotasOma(evaluaciongenerica=campo, manual=False,
                                                                       calificacion=0)
                                    auditorianotas.save()

                response = JsonResponse({'resp': True})
            except Exception as ex:
                transaction.set_rollback(True)
                response = JsonResponse({'resp': False, 'mensaje': ex})
            return HttpResponse(response.content)

        elif action == 'exportarnotarecordacademico':
            try:
                curso = Curso.objects.get(pk=int(encrypt(request.POST['id'])))
                for alumno in curso.inscripcioncurso_set.filter(status=True):
                    for asigInscr in AsignaturaInscripcionCurso.objects.filter(status=True, inscripcioncurso=alumno):
                        if asigInscr.nota >= asigInscr.inscripcioncurso.curso.modeloevaluativo.notaaprobar:
                            asigInscr.cierre_materia_asignada()
                if not curso.cerrado:
                    curso.cerrado = True
                    curso.save(request)
                response = JsonResponse({'resp': True})
            except Exception as ex:
                transaction.set_rollback(True)
                response = JsonResponse({'resp': False, 'mensaje': ex})
            return HttpResponse(response.content)

        elif action == 'cerrar_individual':
            try:
                alumno =InscripcionCurso.objects.get(id=int(encrypt(request.POST['id'])))
                for notasmooc in alumno.curso.notas_de_moodle(alumno.inscripcion.persona):
                    for asigInscr in AsignaturaInscripcionCurso.objects.filter(status=True, inscripcioncurso=alumno):
                        campo = asigInscr.campo(notasmooc[1].upper().strip())
                        if type(notasmooc[0]) is Decimal:
                            if null_to_decimal(campo.valor) != float(notasmooc[0]):
                                actualizar_nota_ofimatica(asigInscr.id, notasmooc[1].upper().strip(), notasmooc[0])
                                auditorianotas = AuditoriaNotasOma(evaluaciongenerica=campo, manual=False,
                                                                   calificacion=notasmooc[0])
                                auditorianotas.save()
                        else:
                            if null_to_decimal(campo.valor) != float(0):
                                actualizar_nota_ofimatica(asigInscr.id, notasmooc[1].upper().strip(), notasmooc[0])
                                auditorianotas = AuditoriaNotasOma(evaluaciongenerica=campo, manual=False,
                                                                   calificacion=0)
                                auditorianotas.save()
                        if asigInscr.nota >= asigInscr.inscripcioncurso.curso.modeloevaluativo.notaaprobar:
                            asigInscr.cierre_materia_asignada()
                res_json = {'error': False}
            except Exception as ex:
                transaction.set_rollback(True)
                res_json = {'error': True, "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)

        elif action == 'reporte_certificado':
            try:
                aprobado = InscripcionCurso.objects.filter(pk=int(encrypt(request.POST['id'])))
                resultados_errores = generarCertificadoOfimatica(aprobado, data)
                if len(resultados_errores) > 0:
                    return JsonResponse({"result": "bad",
                                         "mensaje": u"Problemas al generar el Certificado. " + str(
                                             resultados_errores)})
                else:
                    return JsonResponse({"result": "ok", "mensaje": u"Certificado generado exitosamente"})
            except Exception as ex:
                # messages.error(request, ex)
                return JsonResponse({"result": "bad", "mensaje": u"Problemas al ejecutar el reporte. %s" % ex})

        elif action == 'subirarchivoinscripcion':
            try:
                f = MasivoInscripcionForm(request.POST)
                if f.is_valid():
                    nfile = request.FILES['archivo']
                    archivo = io.BytesIO(nfile.read())
                    workbook = load_workbook(filename=archivo, read_only=False)
                    loes = workbook[workbook.sheetnames[0]]
                    ccedula = 1
                    ccorreo = 3
                    id_carrera = 4
                    curso = Curso.objects.get(pk=int(encrypt(request.POST['id'])))
                    for rowx in range(2, loes.max_row + 1):
                        cedula = loes.cell(row=rowx, column=ccedula).value
                        correo = loes.cell(row=rowx, column=ccorreo).value
                        carrera = loes.cell(row=rowx, column=id_carrera).value
                        insccurso=None
                        if Inscripcion.objects.filter(persona__cedula=cedula, carrera_id=carrera,status=True).exists():
                            inscripcion = Inscripcion.objects.filter(persona__cedula=cedula,carrera_id=carrera, status=True).first()
                            if not correo:
                                correo=inscripcion.persona.emailinst
                            try:
                                insccurso = InscripcionCurso.objects.get(Q(inscripcion=inscripcion) or Q(persona=inscripcion.persona), curso=curso,
                                                                         status=True)
                            except ObjectDoesNotExist:
                                insccurso = InscripcionCurso(inscripcion=inscripcion, curso=curso, correo=correo)
                                insccurso.save(request)
                        # elif Persona.objects.filter(Q(cedula=cedula) or Q(pasaporte=cedula)).exists():
                        #     personainsc=Persona.objects.filter(Q(cedula=cedula) or Q(pasaporte=cedula))
                        #     if not correo:
                        #         correo = personainsc.emailinst
                        #     try:
                        #         insccurso = InscripcionCurso.objects.get(persona=personainsc, curso=curso,
                        #                                                  status=True)
                        #     except ObjectDoesNotExist:
                        #         insccurso = InscripcionCurso(persona=personainsc, curso=curso, correo=correo)
                        #         insccurso.save(request)
                        if insccurso:
                            for x in AsignaturaCurso.objects.filter(curso=curso, status=True):
                                try:
                                    asiginsccurso = AsignaturaInscripcionCurso.objects.get(asignaturacurso=x,
                                                                                           inscripcioncurso=insccurso,
                                                                                           status=True)
                                except ObjectDoesNotExist:
                                    asiginsccurso = AsignaturaInscripcionCurso(asignaturacurso=x,
                                                                               inscripcioncurso=insccurso)
                                    asiginsccurso.save(request)

                                for y in DetalleModeloEvaluativo.objects.filter(modelo=curso.modeloevaluativo,
                                                                                status=True):
                                    try:
                                        evaluagenerica = EvaluacionGenerica.objects.get(detallemodeloevaluativo=y,
                                                                                        asignaturainscripcion=asiginsccurso,
                                                                                        status=True)
                                    except ObjectDoesNotExist:
                                        evaluagenerica = EvaluacionGenerica(detallemodeloevaluativo=y,
                                                                            asignaturainscripcion=asiginsccurso)
                                        evaluagenerica.save(request)
                    return JsonResponse({'result': False, "mensaje": u"Inscripcion exitosa"})
                else:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()],
                                         "mensaje": "Error en el formulario"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, "mensaje": u"Error al inscribir masivo"})

        elif action == 'subirarchivoinscripcionlibre':
            try:
                f = MasivoInscripcionForm(request.POST)
                if f.is_valid():
                    nfile = request.FILES['archivo']
                    archivo = io.BytesIO(nfile.read())
                    workbook = load_workbook(filename=archivo, read_only=False)
                    loes = workbook[workbook.sheetnames[0]]
                    ccedula = 1
                    ccorreo = 3
                    curso = Curso.objects.get(pk=int(encrypt(request.POST['id'])))
                    for rowx in range(2, loes.max_row + 1):
                        cedula = loes.cell(row=rowx, column=ccedula).value
                        correo = loes.cell(row=rowx, column=ccorreo).value
                        if Persona.objects.filter(Q(cedula=cedula) or Q(pasaporte=cedula)).exists():
                            personainsc=Persona.objects.filter(Q(cedula=cedula) or Q(pasaporte=cedula)).first()
                            if not correo:
                                correo = personainsc.email
                            insccurso = None
                            try:
                                insccurso = InscripcionCurso.objects.get(persona=personainsc, curso=curso,
                                                                         status=True)
                            except ObjectDoesNotExist:
                                insccurso = InscripcionCurso(persona=personainsc, curso=curso, correo=correo)
                                insccurso.save(request)
                        for y in DetalleModeloEvaluativo.objects.filter(modelo=curso.modeloevaluativo, status=True):
                            try:
                                evaluagenerica = EvaluacionGenerica.objects.get(detallemodeloevaluativo=y,
                                                                                inscripcioncurso=insccurso,
                                                                                status=True)
                            except ObjectDoesNotExist:
                                evaluagenerica = EvaluacionGenerica(detallemodeloevaluativo=y,
                                                                    inscripcioncurso=insccurso)
                                evaluagenerica.save(request)
                    return JsonResponse({'result': False, "mensaje": u"Inscripcion exitosa"})
                else:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()],
                                         "mensaje": "Error en el formulario"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, "mensaje": u"Error al inscribir masivo"})

        elif action == 'traeralumnosmoodle':
            try:
                curso = Curso.objects.get(pk=request.POST['id'], status=True)
                estudiantes = curso.inscritos()
                primerestudiante = estudiantes.first()
                bandera = True
                modelo_mood = ''
                modelo_sga = ''
                for notasmooc in curso.notas_de_moodle(primerestudiante.matricula.inscripcion.persona):
                    bandera = primerestudiante.evaluacion_generica().filter(detallemodeloevaluativo__nombre=notasmooc[1].upper().strip()).exists()
                    if not bandera:
                        for notasmoocstr in curso.notas_de_moodle(primerestudiante.matricula.inscripcion.persona):
                            modelo_mood += "{}, ".format(notasmoocstr[1])
                        for notassga in primerestudiante.evaluacion_generica():
                            modelo_sga += "{}, ".format(notassga.detallemodeloevaluativo.nombre)
                        return JsonResponse({"result": "bad",
                                             "mensaje": u"Modelo Evaluativo extraido es diferente al modelo existente\nMoodle:\n{}\nSGA:\n{}".format(
                                                 modelo_mood, modelo_sga)})
                listaenviar = estudiantes.values('id','inscripcion__persona__apellido1','inscripcion__persona__apellido2',
                                                                                           'inscripcion__persona__nombres').order_by(
                    'inscripcion__persona__apellido1')
                return JsonResponse({"result": "ok", "cantidad": len(listaenviar), "inscritos": convertir_lista(listaenviar)})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        # elif action == 'traernotaindividual':
        #     try:
        #         curso = Curso.objects.get(pk=request.POST['idmateria'], status=True)
        #         alumno = InscripcionCurso.objects.get(pk=request.POST['id'])
        #         if curso.notas_de_moodle(alumno.matricula.inscripcion.persona):
        #             for notasmooc in curso.notas_de_moodle(alumno.matricula.inscripcion.persona):
        #                 campo = alumno.campo(notasmooc[1].upper().strip())
        #                 if type(notasmooc[0]) is Decimal:
        #                     if null_to_decimal(campo.valor) != float(notasmooc[0]) or (
        #                             alumno.asistenciafinal < campo.detallemodeloevaluativo.modelo.asistenciaaprobar):
        #                         actualizar_nota_planificacion(alumno.id, notasmooc[1].upper().strip(), notasmooc[0])
        #                 else:
        #                     if null_to_decimal(campo.valor) != float(0) or (
        #                             alumno.asistenciafinal < campo.detallemodeloevaluativo.modelo.asistenciaaprobar):
        #                         actualizar_nota_planificacion(alumno.id, notasmooc[1].upper().strip(), notasmooc[0])
        #         else:
        #             for detallemodelo in curso.modeloevaluativo.detallemodeloevaluativo_set.filter(migrarmoodle=True):
        #                 campo = alumno.campo(detallemodelo.nombre)
        #                 actualizar_nota_planificacion(alumno.id, detallemodelo.nombre, 0)
        #         return JsonResponse({"result": "ok"})
        #     except Exception as ex:
        #         import sys
        #         print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
        #         return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos. %s" % ex})


        elif action == 'addmodeloevaluativo':
            try:
                f = ModeloEvaluativoForm(request.POST)
                if f.is_valid():
                    modelo = ModeloEvaluativo(nombre=f.cleaned_data['nombre'],
                                              fecha=datetime.now().date(),
                                              principal=f.cleaned_data['principal'],
                                              activo=f.cleaned_data['activo'],
                                              notamaxima=f.cleaned_data['notamaxima'],
                                              notaaprobar=f.cleaned_data['notaaprobar'],
                                              notarecuperacion=f.cleaned_data['notarecuperacion'],
                                              asistenciaaprobar=f.cleaned_data['asistenciaaprobar'],
                                              asistenciarecuperacion=f.cleaned_data['asistenciarecuperacion'],
                                              notafinaldecimales=f.cleaned_data['notafinaldecimales'],
                                              observaciones=f.cleaned_data['observaciones'])
                    modelo.save(request)
                    if not ModeloEvaluativo.objects.filter(principal=True).exists():
                        modelo.principal = True
                        modelo.save(request)
                    if modelo.principal:
                        for m in ModeloEvaluativo.objects.exclude(id=modelo.id):
                            m.principal = False
                            m.save(request)
                    log(u'Adicionado modelo evaluativo - oma: %s' % modelo, request, "add")
                    return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'editmodeloevaluativo':
            try:
                f = ModeloEvaluativoForm(request.POST)
                if f.is_valid():
                    modelo = ModeloEvaluativo.objects.get(pk=int(encrypt(request.POST['id'])))
                    modelo.nombre = f.cleaned_data['nombre']
                    modelo.fecha = datetime.now().date()
                    modelo.activo = f.cleaned_data['activo']
                    modelo.principal = f.cleaned_data['principal']
                    modelo.notamaxima = f.cleaned_data['notamaxima']
                    modelo.notaaprobar = f.cleaned_data['notaaprobar']
                    modelo.notarecuperacion = f.cleaned_data['notarecuperacion']
                    modelo.asistenciaaprobar = f.cleaned_data['asistenciaaprobar']
                    modelo.asistenciarecuperacion = f.cleaned_data['asistenciarecuperacion']
                    modelo.notafinaldecimales = f.cleaned_data['notafinaldecimales']
                    modelo.observaciones = f.cleaned_data['observaciones']
                    modelo.save(request)
                    if modelo.principal:
                        for m in ModeloEvaluativo.objects.exclude(id=modelo.id):
                            m.principal = False
                            m.save(request)
                    if not ModeloEvaluativo.objects.filter(principal=True).exists():
                        modelo = ModeloEvaluativo.objects.order_by('-fecha')[0]
                        modelo.principal = True
                        modelo.save(request)
                    log(u'Modifico modelo evaluativo: %s' % modelo, request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                     raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'adddetalle':
            try:
                f = DetalleModeloEvaluativoForm(request.POST)
                if f.is_valid():
                    modelo = ModeloEvaluativo.objects.get(pk=int(encrypt(request.POST['id'])))
                    if modelo.detallemodeloevaluativo_set.filter(nombre=f.cleaned_data['nombre']).exists():
                        return JsonResponse({"result": True, "mensaje": u"Ya existe un campo con este nombre."})
                    if modelo.detallemodeloevaluativo_set.filter(orden=f.cleaned_data['orden']).exists():
                        return JsonResponse({"result": True, "mensaje": u"Ya existe un campo con ese numero de orden."})
                    detalle = DetalleModeloEvaluativo(modelo=modelo,
                                                      nombre=f.cleaned_data['nombre'],
                                                      notaminima=f.cleaned_data['notaminima'],
                                                      notamaxima=f.cleaned_data['notamaxima'],
                                                      decimales=f.cleaned_data['decimales'],
                                                      migrarmoodle=f.cleaned_data['migrarmoodle'],
                                                      dependiente=f.cleaned_data['dependiente'],
                                                      dependeasistencia=f.cleaned_data['dependeasistencia'],
                                                      orden=f.cleaned_data['orden'],
                                                      determinaestadofinal=f.cleaned_data['determinaestadofinal'])
                    detalle.save(request)
                    log(u'Adiciono detalle de modelo evaluativo - oma: %s' % detalle, request, "edit")
                    return JsonResponse({"result": False, 'mensaje': 'Registro Exitoso'})
                else:
                     raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": u"Error al guardar los datos."})

        elif action == 'editdetalle':
            try:
                f = DetalleModeloEvaluativoForm(request.POST)
                if f.is_valid():
                    detalle = DetalleModeloEvaluativo.objects.get(pk=int(encrypt(request.POST['id'])))
                    if DetalleModeloEvaluativo.objects.filter(modelo=detalle.modelo, nombre=detalle.nombre).exclude(id=detalle.id).exists():
                        return JsonResponse({"result": "bad", "mensaje": u"Ya existe un campo con ese nombre."})
                    if DetalleModeloEvaluativo.objects.filter(modelo=detalle.modelo, orden=f.cleaned_data['orden']).exclude(id=detalle.id).exists():
                        return JsonResponse({"result": "bad", "mensaje": u"Ya existe un campo con ese numero de orden."})
                    detalle.notaminima = f.cleaned_data['notaminima']
                    detalle.notamaxima = f.cleaned_data['notamaxima']
                    detalle.decimales = f.cleaned_data['decimales']
                    detalle.migrarmoodle = f.cleaned_data['migrarmoodle']
                    detalle.determinaestadofinal = f.cleaned_data['determinaestadofinal']
                    detalle.dependiente = f.cleaned_data['dependiente']
                    detalle.orden = f.cleaned_data['orden']
                    detalle.dependeasistencia = f.cleaned_data['dependeasistencia']
                    detalle.save(request)
                    log(u'Modifico detalle de modelo evaluativo: %s' % detalle, request, "edit")
                    return JsonResponse({"result": False, 'mensaje': 'Edicion Exitosa'})
                else:
                     raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": u"Error al guardar los datos."})

        elif action == 'delmodelo':
            try:
                id=encrypt(request.POST['id'])
                modelo = ModeloEvaluativo.objects.get(pk=int(id))
                log(u"Elimino modelo evaluativo: %s" % modelo, request, "del")
                modelo.status=False
                modelo.save(request)
                if not ModeloEvaluativo.objects.filter(principal=True).exists():
                    modelo = ModeloEvaluativo.objects.order_by('-fecha')[0]
                    modelo.principal = True
                    modelo.save(request)
                return JsonResponse({"error": False})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"error": True, "message": u"Error al eliminar los datos. %s"%(ex.__str__())})

        elif action == 'deldetalle':
            try:
                detalle = DetalleModeloEvaluativo.objects.get(pk=int(encrypt(request.POST['id'])))
                if detalle.en_uso():
                    return JsonResponse({"result": "bad", "mensaje": u"El detalle se encuentra en uso."})
                log(u"Elimino campo de modelo evaluativo - oma: %s" % detalle, request, "del")
                detalle.status = False
                detalle.save()
                res_json = {'error': False}
            except Exception as ex:
                transaction.set_rollback(True)
                res_json = {'error': True, "message": "Error: {}".format(ex)}
                # return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})
            return JsonResponse(res_json, safe=False)

        elif action == 'logica':
            try:
                modelo = ModeloEvaluativo.objects.get(pk=request.POST['id'])
                f = LogicaModeloEvaluativoForm(request.POST)
                if f.is_valid():
                    modelo.logicamodelo = f.cleaned_data['logica']
                    modelo.save(request)
                log(u"Modifico logica calculo: %s" % modelo, request, "del")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'inactivar':
            try:
                modelo = ModeloEvaluativo.objects.get(pk=request.POST['id'])
                if modelo.activo == True:
                    modelo.activo = False
                else:
                    modelo.activo=True
                modelo.save(request)
                log(u"Inactivo Modelo Evaluativo: %s" % modelo, request, "del")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})


        return JsonResponse({"result": True, "mensaje": u"Solicitud Incorrecta."})

    else:
        if 'action' in request.GET:
            data['action'] = action = request.GET['action']

            if action == 'add':
                try:
                    data['title'] = u'Adicionar nuevo Curso'
                    data['action'] = request.GET['action']
                    data['form'] = CursoForm()
                    template = get_template("oma_curso/modal/formCurso.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, 'message': str(ex)})

            elif action == 'addinscripcioncurso':
                try:
                    data['title'] = u'Adicionar nueva Inscripion'
                    data['id'] = id = int(encrypt(request.GET['id']))
                    if AsignaturaCurso.objects.filter(status=True, curso_id=id).exists():
                        data['inscripciones'] = Inscripcion.objects.filter(status=True, activo=True)
                        data['action'] = request.GET['action']
                        form = InscripcionCursoForm()
                        form.fields['inscripcion'].queryset = Inscripcion.objects.none()
                        data['form'] = form
                        template = get_template("oma_curso/modal/formInscripcionCurso.html")
                        return JsonResponse({"result": True, 'data': template.render(data)})
                    else:
                        return JsonResponse({"result": False, 'message': 'Curso sin asignaturas registradas.'})
                except Exception as ex:
                    return JsonResponse({"result": False, 'message': str(ex)})

            elif action == 'addasignaturacurso':
                try:
                    data['title'] = u'Adicionar nueva Asignatura'
                    data['id'] = id = int(encrypt(request.GET['id']))
                    if not InscripcionCurso.objects.filter(status=True,curso_id=id).exists():
                        data['action'] = request.GET['action']
                        form = AsignaturaCursoForm()
                        data['form'] = form
                        template = get_template("oma_curso/modal/formAsignaturaCurso.html")
                        return JsonResponse({"result": True, 'data': template.render(data)})
                    else:
                        return JsonResponse({"result": False, 'message': 'Ya hay inscritos en este curso.'})
                except Exception as ex:
                    return JsonResponse({"result": False, 'message': str(ex)})

            elif action == 'edit':
                try:
                    data['title'] = u'Modificar Curso'
                    data['action'] = request.GET['action']
                    data['id'] = id = int(encrypt(request.GET['id']))
                    data['curso'] = curso = Curso.objects.get(pk=encrypt(request.GET['id']))
                    form = CursoForm(initial={'nombre': curso.nombre,
                                              'codigo': curso.codigo,
                                              'modeloevaluativo': curso.modeloevaluativo,
                                              'fechainicio': curso.fecha_inicio,
                                              'fechafin': curso.fecha_fin,
                                              'horas': curso.horas,
                                              'creditos': curso.creditos,
                                              'enlacegrabacion': curso.enlacegrabacion,
                                              'enlacereuniondocente': curso.enlacereuniondocente,
                                              'enlacepresentacioncurso': curso.enlacepresentacioncurso,
                                              'idcursomoodle': curso.idcursomoodle
                                              })
                    data['form'] = form
                    template = get_template("oma_curso/modal/formCurso.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, 'message': str(ex)})

            elif action == 'buscarinscripcion':
                try:
                    q = request.GET['q'].upper().strip()
                    s = q.split(" ")
                    if len(s) == 1:
                        per = Inscripcion.objects.filter(
                            (Q(persona__nombres__icontains=q) | Q(persona__apellido1__icontains=q) | Q(
                                persona__apellido2__icontains=q) | Q(persona__cedula__contains=q)),
                            Q(status=True, activo=True))[:15]
                    elif len(s) == 2:
                        per = Inscripcion.objects.filter(
                            (Q(persona__apellido1__icontains=s[0]) & Q(persona__apellido2__icontains=s[1])) |
                            (Q(persona__nombres__icontains=s[0]) & Q(persona__nombres__icontains=s[1])) |
                            (Q(persona__nombres__icontains=s[0]) & Q(persona__apellido1__icontains=s[1]))).filter(
                            Q(status=True, activo=True))[:15]
                    data = [{'id': qs.pk, 'text': f"{qs.persona.nombre_completo_inverso()}",
                             'documento': qs.persona.documento(), 'carrera': qs.carrera.nombre,
                             'foto': qs.persona.get_foto()} for qs in per]
                    return HttpResponse(json.dumps({'status': True, 'results': data}))
                except Exception as ex:
                    pass

            elif action == 'subirarchivoinscripcion':
                try:
                    data['id'] = id = int(encrypt(request.GET['id']))
                    if AsignaturaCurso.objects.filter(status=True, curso_id=id).exists():
                        form = MasivoInscripcionForm()
                        data['form'] = form
                        data['action'] = 'subirarchivoinscripcion'
                        template = get_template("oma_curso/modal/formMigrarInscripcionCurso.html")
                        return JsonResponse({"result": True, 'data': template.render(data)})
                    else:
                        return JsonResponse({"result": False, 'message': 'Curso sin asignaturas registradas.'})
                except Exception as ex:
                    return JsonResponse({"result": "bad", "message": u"Error al obtener los datos."})

            elif action == 'subirarchivoinscripcionlibre':
                try:
                    data['id'] = id = int(encrypt(request.GET['id']))
                    form = MasivoInscripcionForm()
                    data['form'] = form
                    data['action'] = 'subirarchivoinscripcionlibre'
                    template = get_template("oma_curso/modal/formMigrarInscripcionCurso.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": "bad", "message": u"Error al obtener los datos."})

            elif action == 'reporte_generar_masivo':
                try:
                    data['curso'] = curso = Curso.objects.get(pk=int(encrypt(request.GET['curso'])))
                    aprobados = InscripcionCurso.objects.filter(status=True, curso=curso).order_by('-id')
                    # , fecharefrendacion__date__lte = datetime.now().date()
                    if aprobados:
                        # pruebas local
                        # if IS_DEBUG and len(graduados) > 2:
                        #     graduados = graduados[:2]
                        resultados_errores = generarCertificadoOfimatica(aprobados, data)
                        if len(resultados_errores) > 0:
                            return JsonResponse({"result": "bad",
                                                 "mensaje": u"Problemas al generar los certificados."})
                        else:
                            return JsonResponse({"result": "ok",
                                                 "mensaje": u"Certificados masivos generados exitosamente."})
                    else:
                        return JsonResponse({"result": "bad",
                                             "mensaje": u"No hay nuevos registros de aprobados para generar certificado."})
                except Exception as ex:
                    # messages.error(request, f'Error {ex} on line {sys.exc_info()[-1].tb_lineno}')
                    return JsonResponse({"result": "bad",
                                         "mensaje": f"Error al obtener los datos {str(ex)}. Error on line {sys.exc_info()[-1].tb_lineno}"})

            elif action == 'crear_cursos_moodle_complexivo_posgrado':
                try:
                    curso = Curso.objects.get(pk=int(encrypt(request.GET['id'])))
                    curso.crear_actualizar_curso_ofimatica()
                    return JsonResponse({"result": True, 'mensaje':'El curso se creo correctamente.'})
                except Exception as ex:
                    return JsonResponse({"result": False, 'mensaje':'Ha ocurrido un error al crear el curso.'})

            elif action == 'enrolar_actualizar_estudiantes_grupos_posgrado':
                try:
                    curso = Curso.objects.get(pk=int(encrypt(request.GET['id'])))
                    from moodle import moodle
                    tipourl=1
                    curso.crear_actualizar_estudiantes_curso_ofimatica(moodle, tipourl)
                    return JsonResponse({"result": True, 'mensaje':'Los maestrantes se enrolaron al curso de  moodle correctamente.'})
                except Exception as ex:
                    return JsonResponse({"result": False, 'mensaje':'Ha ocurrido un error al enrolar a los maestrantes al curso de moodle.'})

            elif action == 'inscritos':
                try:
                    data['title'] = u'Inscripciones Curso de Ofimática'
                    data['id'] =id= int(encrypt(request.GET['id']))
                    data['curso'] = curso = Curso.objects.get(pk=int(encrypt(request.GET['id'])))
                    url_vars, search, filtro = f'&action={action}&id={encrypt(id)}', request.GET.get('s', ''), Q(status=True, curso=curso)
                    if search:
                        q = search.upper().strip()
                        s = q.split(" ")
                        if len(s) == 1:
                            filtro = filtro & (
                                    Q(inscripcion__persona__nombres__icontains=q) |
                                    Q(inscripcion__persona__cedula__icontains=q) |
                                    Q(inscripcion__persona__apellido1__icontains=q) |
                                    Q(inscripcion__persona__apellido2__icontains=q)|
                                    Q(persona__nombres__icontains=q) |
                                     Q(persona__cedula__icontains=q) |
                                     Q(persona__apellido1__icontains=q) |
                                     Q(persona__apellido2__icontains=q)
                            )
                        elif len(s) == 2:
                            filtro = filtro & (
                                    (Q(inscripcion__persona__nombres__icontains=s[0]) & Q(inscripcion__persona__apellido1__icontains=s[1])) |
                                    (Q(inscripcion__persona__apellido1__icontains=s[0]) & Q(inscripcion__persona__apellido2__contains=s[1]))|
                                    (Q(persona__nombres__icontains=s[0]) & Q(persona__apellido1__icontains=s[1])) |
                                    (Q(persona__apellido1__icontains=s[0]) & Q(persona__apellido2__contains=s[1])) )
                        data['search'] = search
                        url_vars += "&s={}".format(search)
                    if 'ids' in request.GET:
                        ids = int(encrypt(request.GET['ids']))
                        url_vars += "&id={}".format(ids)
                        filtro = filtro & Q(pk=ids)
                    inscripcioncurso = InscripcionCurso.objects.filter(filtro).order_by('inscripcion__persona__apellido1').distinct()
                    paging = MiPaginador(inscripcioncurso, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['usuario'] = usuario
                    data['page'] = page
                    data['url_vars'] = url_vars
                    data['listado'] = page.object_list
                    return render(request, "oma_curso/viewInscripcionCurso.html", data)
                except Exception as ex:
                    pass
            elif action == 'asignaturas':
                data['title'] = u'Asignaturas Curso de Ofimática'
                data['id'] = int(encrypt(request.GET['id']))
                data['curso'] = curso = Curso.objects.get(pk=int(encrypt(request.GET['id'])))
                url_vars, search, filtro = f'&action={action}', request.GET.get('s', ''), Q(status=True, curso=curso)
                if search:
                    data['search'] = search
                    url_vars += "&s={}".format(search)
                    filtro = filtro & (Q(asignatura__nombre__icontains=search))
                if 'ids' in request.GET:
                    ids = int(encrypt(request.GET['ids']))
                    url_vars += "&id={}".format(ids)
                    filtro = filtro & Q(pk=ids)
                asignaturacurso = AsignaturaCurso.objects.filter(filtro).order_by('-pk').distinct()
                paging = MiPaginador(asignaturacurso, 20)
                p = 1
                try:
                    paginasesion = 1
                    if 'paginador' in request.session:
                        paginasesion = int(request.session['paginador'])
                    if 'page' in request.GET:
                        p = int(request.GET['page'])
                    else:
                        p = paginasesion
                    try:
                        page = paging.page(p)
                    except:
                        p = 1
                    page = paging.page(p)
                except:
                    page = paging.page(p)
                request.session['paginador'] = p
                data['paging'] = paging
                data['rangospaging'] = paging.rangos_paginado(p)
                data['usuario'] = usuario
                data['page'] = page
                data['url_vars'] = url_vars
                data['listado'] = page.object_list
                return render(request, "oma_curso/viewAsignaturas.html", data)

            elif action == 'notasmoodle':
                try:
                    data['title'] = u'Notas de moodle'
                    data['curso'] = curso = Curso.objects.get(pk=int(encrypt(request.GET['id'])))
                    data['inscritos'] = curso.inscritos()
                    return render(request, "oma_curso/notasmoodle.html", data)
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos. %s"%ex})

            elif action == 'modeloevaluativo':
                try:
                    search, desde, hasta, filtro = request.GET.get('s', ''), request.GET.get('desde',
                                                                                             ''), request.GET.get(
                        'hasta', ''), Q(status=True)
                    url_vars = ""
                    data['title'] = u'Modelos evaluativos'

                    if 's' in request.GET:
                        data['s'] = search = request.GET['s']
                        filtro &= Q(nombre__icontains=search)
                        url_vars += f"&s={search}"

                    if desde:
                        data['desde'] = desde
                        filtro &= Q(fecha__gte=desde)
                        url_vars += '&desde=' + desde

                    if hasta:
                        data['hasta'] = hasta
                        url_vars += "&hasta={}".format(hasta)
                        filtro = filtro & Q(fecha__lte=hasta)

                    modeloevolutivo = ModeloEvaluativo.objects.filter(filtro)
                    paging = MiPaginador(modeloevolutivo, 10)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['listado'] = page.object_list
                    data['url_vars'] = url_vars
                    return render(request, "oma_modeloevaluativo/view.html", data)
                except Exception as ex:
                    num_line = sys.exc_info()[-1].tb_lineno
                    ex_err =f"Error al obtener los datos. {ex}. En la linea {num_line}"
                    return HttpResponseRedirect(f'{request.path}?info={ex_err}')

            elif action == 'addmodeloevaluativo':
                try:
                    # puede_realizar_accion(request, 'sga.puede_modificar_modelos_evaluativos')
                    data['title'] = u'Nuevo modelo evaluativo'
                    data['form'] = ModeloEvaluativoForm()
                    return render(request, "oma_modeloevaluativo/add.html", data)
                except Exception as ex:
                    pass

            elif action == 'adddetalle':
                try:
                    #puede_realizar_accion(request, 'sga.puede_modificar_modelos_evaluativos')
                    data['action'] = action
                    data['title'] = u'Nuevo campo del modelo evaluativo'
                    data['modelo'] = modelo = ModeloEvaluativo.objects.get(pk=int(encrypt(request.GET['id'])))
                    ultimodetalle = null_to_numeric(modelo.detallemodeloevaluativo_set.filter(status=True).aggregate(ultimo=Max('orden'))['ultimo'])
                    data['form'] = DetalleModeloEvaluativoForm(initial={'orden': ultimodetalle + 1})
                    template = get_template("oma_modeloevaluativo/modal/formDetalle.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, 'message': str(ex)})

            elif action == 'detalle':
                try:
                    data['title'] = u'Detalle modelo evaluativo'
                    data['modelo'] = modelo = ModeloEvaluativo.objects.get(pk=int(encrypt(request.GET['id'])))
                    data['campos'] = modelo.detallemodeloevaluativo_set.filter(status=True)
                    return render(request, "oma_modeloevaluativo/detalle.html", data)
                except Exception as ex:
                    pass

            elif action == 'editmodeloevaluativo':
                try:
                    #puede_realizar_accion(request, 'sga.puede_modificar_modelos_evaluativos')
                    data['title'] = u'Editar modelo evaluativo'
                    data['id'] = ids = int(encrypt(request.GET['id']))
                    data['modelo'] = modelo = ModeloEvaluativo.objects.get(pk=ids)
                    data['form'] = form = ModeloEvaluativoForm(initial={'nombre': modelo.nombre,
                                                                        'principal': modelo.principal,
                                                                        'activo': modelo.activo,
                                                                        'notamaxima': modelo.notamaxima,
                                                                        'notaaprobar': modelo.notaaprobar,
                                                                        'notarecuperacion': modelo.notarecuperacion,
                                                                        'asistenciaaprobar': modelo.asistenciaaprobar,
                                                                        'asistenciarecuperacion': modelo.asistenciarecuperacion,
                                                                        'notafinaldecimales': modelo.notafinaldecimales,
                                                                        'observaciones': modelo.observaciones})

                    return render(request, "oma_modeloevaluativo/add.html", data)
                except Exception as ex:
                    pass

            elif action == 'editdetalle':
                try:
                    #puede_realizar_accion(request, 'sga.puede_modificar_modelos_evaluativos')
                    data['title'] = u'Editar campo'
                    data['action'] = request.GET['action']
                    data['detalle'] = detalle = DetalleModeloEvaluativo.objects.get(pk=int(encrypt(request.GET['id'])))
                    form = DetalleModeloEvaluativoForm(initial={'nombre': detalle.nombre,
                                                                'notaminima': detalle.notaminima,
                                                                'notamaxima': detalle.notamaxima,
                                                                'decimales': detalle.decimales,
                                                                'dependiente': detalle.dependiente,
                                                                'determinaestadofinal': detalle.determinaestadofinal,
                                                                'orden': detalle.orden,
                                                                'dependeasistencia': detalle.dependeasistencia})
                    form.editar()
                    data['form'] = form
                    template = get_template("oma_modeloevaluativo/modal/formDetalle.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, 'message': str(ex)})

            elif action == 'logica':
                try:
                    data['title'] = u'Fórmulas matemáticas para el calculo del modelo'
                    data['modelo'] = modelo = ModeloEvaluativo.objects.get(pk=int(encrypt(request.GET['id'])))
                    data['form'] = LogicaModeloEvaluativoForm(initial={'logica': modelo.logicamodelo})
                    return render(request, "oma_modeloevaluativo/logica.html", data)
                except Exception as ex:
                    pass

            elif action == 'delmodelo':
                try:
                    #puede_realizar_accion(request, 'sga.puede_modificar_modelos_evaluativos')
                    data['title'] = u'Eliminar modelo evaluativo'
                    data['modelo'] = ModeloEvaluativo.objects.get(pk=request.GET['id'])
                    return render(request, "adm_modelosevaluativos/delmodelo.html", data)
                except Exception as ex:
                    pass

            elif action == 'verauditoria':
                try:
                    data['title'] = u'Auditoria'
                    alumno = InscripcionCurso.objects.get(id=int(encrypt(request.GET['id'])))
                    modelo = alumno.curso.modeloevaluativo
                    id_asignaturas_alumno=AsignaturaInscripcionCurso.objects.values_list('id',flat=True).filter(status=True,inscripcioncurso=alumno)
                    id_detalles= DetalleModeloEvaluativo.objects.values_list('id',flat=True).filter(status=True,modelo=modelo).order_by('orden')
                    id_evaluaciones= EvaluacionGenerica.objects.values_list('id',flat=True).filter(status=True,
                                                                                                   asignaturainscripcion_id__in=id_asignaturas_alumno,
                                                                                                   detallemodeloevaluativo_id__in=id_detalles)
                    data['auditorianotas'] = AuditoriaNotasOma.objects.filter(status=True, evaluaciongenerica_id__in=id_evaluaciones).order_by('evaluaciongenerica__asignaturainscripcion__asignaturacurso__asignatura__nombre')
                    template = get_template("oma_curso/modal/verauditoria.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, 'message': str(ex)})

            return HttpResponseRedirect(request.path)

        else:
            try:
                data['title'] = u'Cursos de Ofimática'
                url_vars, search, filtro = '', request.GET.get('s', ''), Q(status=True)
                if search:
                    data['search'] = search
                    url_vars += "&s={}".format(search)
                    filtro = filtro & (Q(nombre__icontains=search)|Q(codigo__icontains=search))
                if 'id' in request.GET:
                    ids = int(request.GET['id'])
                    url_vars += "&id={}".format(ids)
                    filtro = filtro & Q(pk=ids)
                cursos = Curso.objects.filter(filtro).order_by('-pk').distinct()
                paging = MiPaginador(cursos, 20)
                p = 1
                try:
                    paginasesion = 1
                    if 'paginador' in request.session:
                        paginasesion = int(request.session['paginador'])
                    if 'page' in request.GET:
                        p = int(request.GET['page'])
                    else:
                        p = paginasesion
                    try:
                        page = paging.page(p)
                    except:
                        p = 1
                    page = paging.page(p)
                except:
                    page = paging.page(p)
                request.session['paginador'] = p
                data['paging'] = paging
                data['rangospaging'] = paging.rangos_paginado(p)
                data['usuario'] = usuario
                data['page'] = page
                data['url_vars'] = url_vars
                data['listado'] = page.object_list
                return render(request, "oma_curso/view.html", data)
            except Exception as ex:
                pass

def generarCertificadoOfimatica(aprobados, data, IS_DEBUG=False):
    dominio_sistema = 'http://127.0.0.1:8000'
    if not IS_DEBUG:
        dominio_sistema = 'https://sga.unemi.edu.ec'
    data["DOMINIO_DEL_SISTEMA"] = dominio_sistema
    lista_correctos = []
    lista_errores = []
    for aprobado in aprobados:
    # with transaction.atomic():
        try:
            if aprobado.asignaturainscripcioncurso_set.filter(status=True, estado=1).exists():
                data['aprobado'] = aprobado
                temp = lambda x: remover_caracteres_tildes_unicode(remover_caracteres_especiales_unicode(x.__str__()))
                data['curso'] = curso = aprobado.curso
                data['fecha_curso'] =fecha_curso =  fecha_certificado_curso(curso.fecha_inicio, curso.fecha_fin)
                data['hoy'] = curso.fecha_fin  + timedelta(days=1)

                qrname = 'qr_certificado_oma_' + str(aprobado.id)
                folder = os.path.join(os.path.join(SITE_STORAGE, 'media', 'qrcode', 'omaCertificados', 'qr'))
                directory = os.path.join(os.path.join(SITE_STORAGE, 'media', 'qrcode', 'omaCertificados'))
                os.makedirs(f'{directory}/qr/', exist_ok=True)
                try:
                    os.stat(directory)
                except:
                    os.mkdir(directory)
                nombrepersona = temp(aprobado.inscripcion.persona.__str__()).replace(' ', '_')
                htmlname = 'CERTIFICADO_OFIMATICA_{}_{}'.format(nombrepersona, random.randint(1, 100000).__str__())
                urlname = "/media/qrcode/omaCertificados/%s" % htmlname
                # rutahtml = SITE_STORAGE + urlname
                data['url_qr'] = url_qr = f'{SITE_STORAGE}/media/qrcode/omaCertificados/qr/{htmlname}.png'
                # if os.path.isfile(rutahtml):
                #     os.remove(rutahtml)
                url = pyqrcode.create(f'{dominio_sistema}/media/qrcode/omaCertificados/{htmlname}.pdf')
                imageqr = url.png(f'{directory}/qr/{htmlname}.png', 16, '#000000')
                data['qrname'] = 'qr' + qrname
                valida = conviert_html_to_pdfsaveqr_omacertificado(
                    'oma_curso/certificado/formato_certificado.html',
                    {'pagesize': 'A4', 'data': data},
                    htmlname + '.pdf'
                )
                if valida:
                    aprobado.archivocertificado = 'qrcode/omaCertificados/' + htmlname + '.pdf'
                    aprobado.save()

                    tituloemail = "Generación de Certificado"

                    send_html_mail(tituloemail,
                                   "emails/confirmacion_generacion_certificado_oma.html",
                                   {'sistema': u'EPUNEMI',
                                    'saludo': 'Estimada' if aprobado.inscripcion.persona.sexo.id == 1 else 'Estimado',
                                    'aprobado': aprobado,
                                    'curso': aprobado.curso
                                    },
                                   aprobado.lista_emails_envio_oma(),
                                   [],
                                   cuenta=CUENTAS_CORREOS[16][1]
                                   )

                    lista_correctos.append(f'{aprobado.inscripcion.persona.cedula} [{aprobado.id}]\n')
            elif "python".lower() in aprobado.curso.nombre.lower():
                data['aprobado'] = aprobado
                temp = lambda x: remover_caracteres_tildes_unicode(remover_caracteres_especiales_unicode(x.__str__()))
                data['curso'] = curso = aprobado.curso
                data['fecha_curso'] = fecha_curso = fecha_certificado_curso(curso.fecha_inicio, curso.fecha_fin)
                data['hoy'] = curso.fecha_fin + timedelta(days=1)

                qrname = 'qr_certificado_oma_' + str(aprobado.id)
                folder = os.path.join(os.path.join(SITE_STORAGE, 'media', 'qrcode', 'omaCertificados', 'qr'))
                directory = os.path.join(os.path.join(SITE_STORAGE, 'media', 'qrcode', 'omaCertificados'))
                os.makedirs(f'{directory}/qr/', exist_ok=True)
                try:
                    os.stat(directory)
                except:
                    os.mkdir(directory)
                nombrepersona = temp(aprobado.persona.__str__()).replace(' ', '_')
                htmlname = 'CERTIFICADO_PYTHON_{}_{}'.format(nombrepersona, random.randint(1, 100000).__str__())
                urlname = "/media/qrcode/omaCertificados/%s" % htmlname
                # rutahtml = SITE_STORAGE + urlname
                data['url_qr'] = url_qr = f'{SITE_STORAGE}/media/qrcode/omaCertificados/qr/{htmlname}.png'
                # if os.path.isfile(rutahtml):
                #     os.remove(rutahtml)
                url = pyqrcode.create(f'{dominio_sistema}/media/qrcode/omaCertificados/{htmlname}.pdf')
                imageqr = url.png(f'{directory}/qr/{htmlname}.png', 16, '#000000')
                data['qrname'] = 'qr' + qrname
                valida = conviert_html_to_pdfsaveqr_omacertificado(
                    'oma_curso/certificado/formato_certificado_new.html',
                    {'pagesize': 'A4', 'data': data},
                    htmlname + '.pdf'
                )
                if valida:
                    aprobado.archivocertificado = 'qrcode/omaCertificados/' + htmlname + '.pdf'
                    aprobado.save()

                    tituloemail = "Generación de Certificado"

                    send_html_mail(tituloemail,
                                   "emails/confirmacion_generacion_certificado_oma.html",
                                   {'sistema': u'EPUNEMI',
                                    'saludo': 'Estimada' if aprobado.persona.sexo.id == 1 else 'Estimado',
                                    'aprobado': aprobado,
                                    'curso': aprobado.curso
                                    },
                                   aprobado.lista_emails_envio_oma(),
                                   [],
                                   cuenta=CUENTAS_CORREOS[16][1]
                                   )

                    lista_correctos.append(f'{aprobado.persona.cedula} [{aprobado.id}]\n')
        except Exception as ex:
            lista_errores.append(
            f'{aprobado.inscripcion.persona.cedula if aprobado.inscripcion else aprobado.persona.cedula } [{aprobado.id}] error {str(ex)} on line {str(sys.exc_info()[-1].tb_lineno)}\n')
    return lista_errores

def fecha_certificado_curso(fecha1,fecha2):
    mes = ["enero", "febrero", "marzo", "abril", "mayo", "junio", "julio", "agosto", "septiembre",
                "octubre", "noviembre", "diciembre"]

    if fecha1.month == fecha2.month:
        fecha = '{} al {} de {} de {}'.format(fecha1.day, fecha2.day, mes[fecha2.month - 1], fecha2.year)
    elif fecha1.year == fecha2.year:
        fecha = '{} de {} al {} de {} de {}'.format(fecha1.day, mes[fecha1.month-1], fecha2.day, mes[fecha2.month-1], fecha2.year)
    else:
        fecha = '{} de {} de {} al {} de {} de {}'.format(fecha1.day, mes[fecha1.month-1], fecha1.year, fecha2.day, mes[fecha2.month-1], fecha2.year)
    return fecha

def actualizar_nota_ofimatica(asigInscr_id, sel_id, valor):
    asignaturainscripcion = AsignaturaInscripcionCurso.objects.get(pk=asigInscr_id)
    sel = sel_id
    datos = {"result": "ok"}
    modeloevaluativo = asignaturainscripcion.inscripcioncurso.curso.modeloevaluativo
    campomodelo = modeloevaluativo.campo(sel)
    try:
        if not valor:
            valor = null_to_decimal(float(valor), campomodelo.decimales)
        if valor >= campomodelo.notamaxima:
            valor = campomodelo.notamaxima
        elif valor <= campomodelo.notaminima:
            valor = campomodelo.notaminima
    except:
        valor = campomodelo.notaminima
    campo = asignaturainscripcion.campo(sel)
    campo.valor = valor
    campo.save()
    # FUNCION DIMAMICA
    d = locals()
    exec(modeloevaluativo.logicamodelo, globals(), d)
    d['calculo_modelo_evaluativo'](asignaturainscripcion)
    # calculo_modelo_evaluativo(materiaasignada)
    asignaturainscripcion.nota = null_to_decimal(asignaturainscripcion.nota, modeloevaluativo.notafinaldecimales)
    if asignaturainscripcion.nota > modeloevaluativo.notamaxima:
        asignaturainscripcion.nota = modeloevaluativo.notamaxima
    asignaturainscripcion.save()
    asignaturainscripcion.actualiza_estado()
