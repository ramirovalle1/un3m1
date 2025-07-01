#!/usr/bin/env python
# -*- coding: utf-8 -*-
import os
import sys
import openpyxl

# SITE_ROOT = os.path.dirname(os.path.realpath(__file__))
YOUR_PATH = os.path.dirname(os.path.realpath(__file__))
# print(f"YOUR_PATH: {YOUR_PATH}")
SITE_ROOT = os.path.dirname(os.path.dirname(YOUR_PATH))
SITE_ROOT = os.path.join(SITE_ROOT, '')
# print(f"SITE_ROOT: {SITE_ROOT}")
your_djangoproject_home = os.path.split(SITE_ROOT)[0]
# print(f"your_djangoproject_home: {your_djangoproject_home}")
sys.path.append(your_djangoproject_home)

from django.core.wsgi import get_wsgi_application
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "settings")
application = get_wsgi_application()
from sga.models import Persona, Notificacion
from sagest.models import CapInscritoIpec, Rubro, CapEventoPeriodoIpec, CuentaContable
from datetime import datetime, timedelta
import warnings
import csv
from django.db import transaction
from sagest.adm_capacitacioneventoperiodoipec import buscarIdPersonaEpunemi, buscarIdTipootrorubroEpunemi, buscarPagosEpunemiRubroUnemi
from sga.funciones import salvaRubrosEpunemiEdcon
from django.db import transaction, connections
warnings.filterwarnings('ignore', message='Unverified HTTPS request')

def salvaRubrosUnemi(model, action, qs_nuevo=None, qs_anterior=None):
    from sagest.models import logRubros
    import json
    arr = []
    anterior = {}
    nuevo = {}
    if qs_anterior:
        anterior["fields"] = {}
        for x in qs_anterior:
            for k, v in x.items():
                anterior["fields"][k] = str(v)
        anterior["pk"] = qs_anterior[0]["id"]
        anterior["fields"]["__ff_detalle_ff__"] = "ANTERIOR"
        anterior["model"] = "{}.{}".format(model._meta.app_label, model._meta.model_name)
        arr.append(anterior)
    if qs_nuevo:
        nuevo["fields"] = {}
        for x in qs_nuevo:
            for k, v in x.items():
                nuevo["fields"][k] = str(v)
        nuevo["pk"] = qs_nuevo[0]["id"]
        nuevo["fields"]["__ff_detalle_ff__"] = "NUEVO"
        nuevo["model"] = "{}.{}".format(model._meta.app_label, model._meta.model_name)
        arr.append(nuevo)
    data_json = json.dumps(arr)
    userid = 1
    auditusuariotabla = logRubros(usuario_id=userid,
                                  idrubro=model.id,
                                  rubroname=model.nombre,
                                  cedulapersona=model.persona.cedula,
                                  persona=model.persona.__str__(),
                                  accion=action.upper(),
                                  datos_json=data_json)
    auditusuariotabla.save()

# Sincronizr rubros de unemi a epunemi edcon
print(u"Inicio sincronizar rubros edcon")
try:
    # capturar eventos ipec
    ideventos = CapEventoPeriodoIpec.objects.values_list('id', flat= True).filter(id__in=(243, 256, 250, 252, 257, 260, 258, 259, 254, 255))
    # ideventos = CapEventoPeriodoIpec.objects.values_list('id', flat= True).filter(id__in=(256, 258, 257, 254))
    mensaje_notificar = ''
    for idevento in ideventos:
        # rubros unemi
        inscritos_activos = CapInscritoIpec.objects.filter(status=True, desactivado=False, capeventoperiodo_id=idevento).values_list('participante_id')
        rubrosunemi = Rubro.objects.filter(status=True, capeventoperiodoipec__id=idevento, persona_id__in=inscritos_activos)
        # solo inscritos NO desactivados rubros unemi a migrar
        # incio de migracion epunemi
        action = 'MIGRADO_DESDE_UNEMI'
        # Valida que rubrosunemi sea una lista o queryset

            # Valida que lista o queryset no esté vacía

        rubrospagados, total, lista_rubros_unemi, lista_rubros_epunemi = [], 0, [], []
        for r in rubrosunemi:
            lista_rubros_unemi.append(r.persona.cedula)
            with transaction.atomic():
                try:
                    # idperunemi = rubrounemi.persona.id
                    identificacion = r.persona.identificacion()

                    # rubrosunemi = RubrosBuscar(idperunemi)
                    # Buscar ID persona EPUNEMI
                    idpersonaepunemi = buscarIdPersonaEpunemi(identificacion)
                    cursor = connections['epunemi'].cursor()

                    if idpersonaepunemi is None:
                        personaunemi = r.persona
                        sql = """ INSERT INTO sga_persona (status, nombres, apellido1, apellido2, cedula, ruc, pasaporte, 
                                            nacimiento, tipopersona, sector, direccion,  direccion2, num_direccion, telefono, telefono_conv, email,
                                            contribuyenteespecial, anioresidencia, nacionalidad, ciudad, referencia, emailinst,
                                            identificacioninstitucion, regitrocertificacion, libretamilitar, servidorcarrera, concursomeritos, 
                                            telefonoextension, tipocelular, periodosabatico, real, lgtbi, datosactualizados, 
                                            confirmarextensiontelefonia, acumuladecimo, acumulafondoreserva, representantelegal, inscripcioncurso, 
                                            unemi, idunemi)
                                            VALUES(TRUE, '%s', '%s', '%s', '%s', '%s', '%s', '/%s/', %s, '%s', '%s', '%s', '%s', '%s', '%s', '%s', 
                                            FALSE, 0, '', '', '', '', '', '', '', FALSE, FALSE, '', 0, FALSE, TRUE, FALSE, 0, FALSE, TRUE, FALSE, FALSE, FALSE, FALSE, 0); """ % (
                            personaunemi.nombres,
                            personaunemi.apellido1,
                            personaunemi.apellido2,
                            personaunemi.cedula,
                            personaunemi.ruc if personaunemi.ruc else '',
                            personaunemi.pasaporte if personaunemi.pasaporte else '',
                            personaunemi.nacimiento,
                            personaunemi.tipopersona if personaunemi.tipopersona else 1,
                            personaunemi.sector if personaunemi.sector else '',
                            personaunemi.direccion if personaunemi.direccion else '',
                            personaunemi.direccion2 if personaunemi.direccion2 else '',
                            personaunemi.num_direccion if personaunemi.num_direccion else '',
                            personaunemi.telefono if personaunemi.telefono else '',
                            personaunemi.telefono_conv if personaunemi.telefono_conv else '',
                            personaunemi.email if personaunemi.email else '')
                        cursor.execute(sql)
                        # Actualizar persona
                        # Verificar que exista en epunemi el id de estos datos
                        sexoepunemi, parroquiaepunemi, cantonepunemi, provinciaepunemi, paisepunemi = None, None, None, None, None

                        # Buscar los datos en EPUNEMI
                        if personaunemi.sexo:
                            sql = """SELECT sexo.id FROM sga_sexo AS sexo WHERE sexo.id='%s'  AND sexo.status=TRUE;  """ % (
                                personaunemi.sexo.id)
                            cursor.execute(sql)
                            sexoepunemi = cursor.fetchone()
                            if sexoepunemi:
                                sql = """UPDATE sga_persona SET sexo_id='%s' WHERE cedula='%s'; """ % (
                                    sexoepunemi[0], personaunemi.cedula)
                                cursor.execute(sql)

                        if personaunemi.parroquia:
                            sql = """SELECT pa.id FROM sga_parroquia AS pa WHERE pa.id='%s'  AND pa.status=TRUE;  """ % (
                                personaunemi.parroquia.id)
                            cursor.execute(sql)
                            parroquiaepunemi = cursor.fetchone()

                            if parroquiaepunemi:
                                sql = """UPDATE sga_persona SET parroquia_id='%s' WHERE cedula='%s'; """ % (
                                    parroquiaepunemi[0], personaunemi.cedula)
                                cursor.execute(sql)

                        if personaunemi.canton:
                            sql = """SELECT ca.id FROM sga_canton AS ca WHERE ca.id='%s'  AND ca.status=TRUE;  """ % (
                                personaunemi.canton.id)
                            cursor.execute(sql)
                            cantonepunemi = cursor.fetchone()

                            if cantonepunemi:
                                sql = """UPDATE sga_persona SET canton_id='%s' WHERE cedula='%s'; """ % (
                                    cantonepunemi[0], personaunemi.cedula)
                                cursor.execute(sql)

                        if personaunemi.provincia:
                            sql = """SELECT pro.id FROM sga_provincia AS pro WHERE pro.id='%s'  AND pro.status=TRUE;  """ % (
                                personaunemi.provincia.id)
                            cursor.execute(sql)
                            provinciaepunemi = cursor.fetchone()

                            if provinciaepunemi:
                                sql = """UPDATE sga_persona SET provincia_id='%s' WHERE cedula='%s'; """ % (
                                    provinciaepunemi[0], personaunemi.cedula)
                                cursor.execute(sql)

                        if personaunemi.pais:
                            sql = """SELECT pai.id FROM sga_pais AS pai WHERE pai.id='%s'  AND pai.status=TRUE;  """ % (
                                personaunemi.pais.id)
                            cursor.execute(sql)
                            paisepunemi = cursor.fetchone()

                            if paisepunemi:
                                sql = """UPDATE sga_persona SET pais_id='%s' WHERE cedula='%s'; """ % (
                                    paisepunemi[0], personaunemi.cedula)
                                cursor.execute(sql)

                        # ID DE PERSONA EPUNEMI creado
                        idpersonaepunemi = buscarIdPersonaEpunemi(identificacion)

                    idpersonaepunemi = idpersonaepunemi[0]

                    # Consulto id de Tipo otro rubro de EPUNEMI
                    idtipootrorubroepunemi = buscarIdTipootrorubroEpunemi(r.tipo.id)

                    if not idtipootrorubroepunemi:
                        tipootrorubrounemi = r.tipo

                        # Consulto id Centro costo de EPUNEMI
                        sql = """SELECT id FROM sagest_centrocosto WHERE status=True AND unemi=True AND tipo=%s;""" % (
                            tipootrorubrounemi.tiporubro)
                        cursor.execute(sql)
                        centrocostoepunemi = cursor.fetchone()
                        idcentrocostoepunemi = centrocostoepunemi[0]

                        # Valido que no se guarde duplicado en Tipo otro rubro EPUNEMI
                        sql = """SELECT id FROM sagest_tipootrorubro WHERE status=True AND nombre ILIKE '%s'; """ % (
                            tipootrorubrounemi.nombre)
                        cursor.execute(sql)
                        idtipootrorubroepunemi = cursor.fetchone()

                        if not idtipootrorubroepunemi:
                            # Consulto id y partida_id de la Cuenta contable de EPUNEMI
                            cuentacontable = CuentaContable.objects.filter(status=True, partida_id=int(
                                tipootrorubrounemi.partida_id)).first()
                            sql = """SELECT id, partida_id FROM sagest_cuentacontable WHERE id=%s; """ % (
                                cuentacontable.id)
                            cursor.execute(sql)
                            cuentacontableepunemi = cursor.fetchone()

                            # Creo el tipo otro rubro en epunemi
                            sql = """ Insert Into sagest_tipootrorubro (status, nombre, partida_id, valor, interface, activo, 
                            ivaaplicado_id, nofactura, exportabanco, cuentacontable_id, centrocosto_id, tiporubro, 
                            idtipootrorubrounemi, unemi, es_especie, es_convalidacionconocimiento)
                            VALUES(TRUE, '%s', %s, %s, FALSE, TRUE, %s, FALSE, TRUE, %s, %s, 1, %s, TRUE, FALSE, FALSE); """ % (
                                r.tipo.nombre,
                                cuentacontableepunemi[1],
                                r.tipo.valor,
                                r.tipo.ivaaplicado.id,
                                cuentacontableepunemi[0],
                                idcentrocostoepunemi,
                                r.tipo.id)
                            cursor.execute(sql)
                            print("*** Tipo otro rubro creado en EPUNEMI ***")
                            # Excelente
                            # Consulto id de Tipo otro rubro de EPUNEMI creado
                            idtipootrorubroepunemi = buscarIdTipootrorubroEpunemi(r.tipo.id)

                    idtipootrorubroepunemi = idtipootrorubroepunemi[0]

                    # Consulto id de Rubro de EPUNEMI
                    # rubroepunemi = buscarIdRubroEpunemi(r.id)
                    sql = """SELECT id FROM sagest_rubro WHERE idrubrounemi=%s AND status=TRUE; """ % (r.id)
                    cursor.execute(sql)
                    rubroepunemi = cursor.fetchone()

                    # Pregunto si el rubro no existe en EPUNEMI
                    if not rubroepunemi:
                        # Creo el rubro en EPUNEMI
                        sql = """ INSERT INTO sagest_rubro (status, persona_id, nombre, cuota, tipocuota, fecha, fechavence,
                            valor, saldo, iva_id, valoriva, totalunemi, valortotal, cancelado, observacion, idrubrounemi, 
                            tipo_id, 
                            fecha_creacion, usuario_creacion_id, tienenotacredito, valornotacredito, valordescuento, 
                            anulado, compromisopago, refinanciado, bloqueado, bloqueadopornovedad, titularcambiado, coactiva)
                            VALUES (TRUE, %s, '%s', %s, %s, '/%s/', '/%s/', %s, %s, %s, %s, %s, %s, %s, '%s', %s, %s, 
                            NOW(), 1, FALSE, 0, 0, FALSE, %s, %s, %s, FALSE, FALSE, %s); """ % (
                            idpersonaepunemi,
                            r.nombre,
                            r.cuota,
                            r.tipocuota,
                            r.fecha,
                            r.fechavence,
                            r.valor,
                            r.saldo,
                            r.iva_id,
                            r.valoriva,
                            r.valor,
                            r.valortotal,
                            r.cancelado,
                            r.observacion,
                            r.id,
                            idtipootrorubroepunemi,
                            r.compromisopago if r.compromisopago else 0,
                            r.refinanciado,
                            r.bloqueado,
                            r.coactiva)
                        cursor.execute(sql)

                        sql = """SELECT row_to_json(r) FROM (SELECT * FROM sagest_rubro WHERE idrubrounemi=%s AND status=TRUE AND anulado=FALSE) r;"""
                        cursor.execute(sql, [r.id])
                        rubroepuneminoanu_dic = cursor.fetchone()
                        rubroepuneminoanu_dic = rubroepuneminoanu_dic[0]

                        # Vincular rubro de unemi con rubro migrado a epunemi
                        r.idrubroepunemi = rubroepuneminoanu_dic['id']  # Guardo solo el id
                        r.save()

                        # guardar auditoría en UNEMI el log del rubro migrado a EPUNEMI
                        qs_nuevo = [vars(r)]
                        salvaRubrosUnemi(r, action, qs_nuevo=qs_nuevo)

                        # guardar auditoría en EPUNEMI el log del rubro migrado desde UNEMI
                        qs_nuevoepunemi = [rubroepuneminoanu_dic]
                        salvaRubrosEpunemiEdcon(rubroepuneminoanu_dic, action, qs_nuevo=qs_nuevoepunemi)
                        total += 1
                        lista_rubros_epunemi.append(r.persona.cedula)
                except Exception as ex:
                    transaction.set_rollback(True)
                    print(f'Error: {ex} Line: {sys.exc_info()[-1].tb_lineno}')
        cursor.close()
        print(f'Se sincronizó en epunemi {total} rubros del evento {idevento} Rubros UNEMI: {lista_rubros_unemi} Rubros EPUNEMI: {lista_rubros_epunemi}, ')
        mensaje_notificar += f'Se sincronizó en epunemi {total} rubros del evento {idevento} Rubros UNEMI: {lista_rubros_unemi} Rubros EPUNEMI: {lista_rubros_epunemi}, '


        # Fin de migracion epunemi
        # Todos los rubros en status false para eliminar a epunemi en caso existan y no se modifica status en unemi
        # rubros_eliminar = Rubro.objects.filter(status=False, capeventoperiodo_id=ideventos)
    notificacion = Notificacion(
        titulo='Sincronizacion masiva de rubros edcon',
        cuerpo=mensaje_notificar,
        destinatario=Persona.objects.get(pk=21966),
        url=f"/adm_capeventoperiodoipec?action=planificacion&id=11",
        content_type=None,
        object_id=None,
        prioridad=1,
        app_label='sga',
        fecha_hora_visible=datetime.now() + timedelta(days=3))
    notificacion.save()
except Exception as ex:
    print(f'Error: {ex} Line: {sys.exc_info()[-1].tb_lineno}')
print(u"FIN sincronizar rubros edcon")

# mar 25 04 2023
# print(u"Inicio masivo subnovedades")
# with transaction.atomic():
#     try:
#         # miarchivo = openpyxl.load_workbook("subnovedades_masivo.csv")
#         # lista = miarchivo.get_sheet_by_name('subnovedades_masivo')
#         # subnovedad = SubnovedadPeriodoRol(archivo=miarchivo)
#         # subnovedad.save()
#         datareader = csv.reader(open('subnovedades_masivo.csv', "rU", encoding='iso-8859-1'), delimiter=';')
#         linea = 1
#         totalsubnov = 0
#         for cols in datareader:
#             a = linea
#             if linea >= 3:
#                 periodorol = int(cols[0].strip())
#                 rubrorol = int(cols[1].strip())
#                 if Persona.objects.filter((Q(perfilusuario__administrativo__isnull=False) | Q(
#                         perfilusuario__profesor__isnull=False)), cedula__icontains=cols[2].strip()).exists():
#                     personarol = Persona.objects.filter((Q(perfilusuario__administrativo__isnull=False) | Q(
#                         perfilusuario__profesor__isnull=False)),
#                                                         cedula__icontains=cols[2].strip())[0]
#                 elif Persona.objects.filter((Q(perfilusuario__administrativo__isnull=False) | Q(
#                         perfilusuario__profesor__isnull=False)), pasaporte__icontains=cols[2].strip()).exists():
#                     personarol = Persona.objects.filter((Q(perfilusuario__administrativo__isnull=False) | Q(
#                         perfilusuario__profesor__isnull=False)),
#                                                         pasaporte__icontains=cols[2].strip())[0]
#                 else:
#                     transaction.set_rollback(True)
#                     print(f'Error numero de identificacion {cols[2].strip()} no existe')
#                 if DetallePeriodoRol.objects.filter(periodo_id=periodorol, persona=personarol, rubro_id=rubrorol,
#                                                     status=True).exists():
#                     detalleperiodorol = DetallePeriodoRol.objects.get(periodo_id=periodorol, persona=personarol,
#                                                                       rubro_id=rubrorol, status=True)
#                 else:
#                     transaction.set_rollback(True)
#                     print(f'Error. En el periodo {periodorol}, la persona {cols[2].strip()} no posee el rubro {rubrorol}.')
#                 detallesubnovedad = DetalleSubnovedadPeriodoRol(
#                     detalleperiodorol=detalleperiodorol,
#                     descripcion=cols[3].strip(),
#                     valor=float(cols[4].strip().replace(',', '.')))
#                 detallesubnovedad.save()
#                 totalsubnov += 1
#             linea += 1
#         print(f'Masivo exitoso, total: {str(totalsubnov)}')
#     except Exception as ex:
#         transaction.set_rollback(True)
#         print(f'Error: {ex} Line: {sys.exc_info()[-1].tb_lineno}')
# print(u"Fin masivo subnovedades")
# #########

#solo posgrado graduados
# graduados = Graduado.objects.filter(status=True, inscripcion__carrera__coordinacion__id=7)
# listaerror = []
# actualizados = 0
# try:
#     # miarchivo = openpyxl.load_workbook("INSIGNIAS_graduadosposgrado_frefrendacion.xlsx")
#     # lista = miarchivo.get_sheet_by_name('graduadosposgrado_hoja')
#     miarchivo = openpyxl.load_workbook("fecha_refrendacion_modificar_titulos_insignia.xlsx")
#     lista = miarchivo.get_sheet_by_name('titulos_posgrado')
#     # lista = miarchivo.wb['graduadosposgrado_hoja']
#     totallista = lista.rows
#     a=0
#     c=0
#     for filas in totallista:
#         try:
#             a += 1
#             if a > 1:
#                 id = str(filas[0].value).strip() if filas[0].value else None
#                 graduado = graduados.get(pk=id)
#                 # if graduado.pk == 12457:
#                 #     print('ok')
#                 if graduado:
#                     fechacelda = str(filas[12].value).strip()
#                     if len(fechacelda)>0:
#                         if len(fechacelda) > 10:
#                             fechacelda = fechacelda[:len(fechacelda)-9]
#                         if '/' in fechacelda:
#                             fecha = datetime.strptime(fechacelda, '%d/%m/%Y').date()
#                         else:
#                             if len(str(fechacelda).split('-').__getitem__(0)) == 4:
#                                 fecha = datetime.strptime(fechacelda, '%Y-%m-%d').date()
#                             else:
#                                 fecha = datetime.strptime(fechacelda, '%d-%m-%Y').date()
#                         if fecha:
#                             # graduado.fecharefrendacion = date(int(fecha[0:4]), int(fecha[5:7]), int(fecha[8:10]))
#                             graduado.fecharefrendacion = date(fecha.year, fecha.month, fecha.day)
#                             graduado.save()
#                             actualizados = actualizados + 1
#         except Exception as ex:
#             print(f'error graduado {a}:, id: {id}, ex:{ex}')
#             listaerror.append(f'id:{id} - error:{str(ex)}')
# except Exception as ex:
#     print('error: %s' % ex)
# print('lista error', str(listaerror), ' total error: ', str(len(listaerror)))
# print('total actualizados', str(actualizados))
#
# # notificacion al administrativo en sga
# if len(listaerror) == 0:
#     titulonotificacion = f"Actualización exitosa"
#     cuerponotificacion = u"Se generó correctamente el proceso. \nTotal actualizados: %s. Errores: %s" % (str(actualizados), str(listaerror))
# else:
#     titulonotificacion = f"Error en el proceso de actualización"
#     cuerponotificacion = u"No se actualizaron %s registros: \n%s, \nTotal correctos: %s." % (
#         str(len(listaerror)), str(listaerror), str(actualizados))
# notificacion = Notificacion(
#     titulo=titulonotificacion,
#     cuerpo=cuerponotificacion,
#     destinatario=Persona.objects.get(pk=21966),
#     url=f"/graduados",
#     content_type=None,
#     object_id=None,
#     prioridad=1,
#     app_label='SGA',
#     fecha_hora_visible=datetime.now() + timedelta(days=3))
# notificacion.save()
# # fin notificacion al administrativo sga
# print("Finaliza")

# --------------
# #!/usr/bin/env python
# # -*- coding: utf-8 -*-
# import os
# import sys
# import openpyxl
#
# # SITE_ROOT = os.path.dirname(os.path.realpath(__file__))
# YOUR_PATH = os.path.dirname(os.path.realpath(__file__))
# # print(f"YOUR_PATH: {YOUR_PATH}")
# SITE_ROOT = os.path.dirname(os.path.dirname(YOUR_PATH))
# SITE_ROOT = os.path.join(SITE_ROOT, '')
# # print(f"SITE_ROOT: {SITE_ROOT}")
# your_djangoproject_home = os.path.split(SITE_ROOT)[0]
# # print(f"your_djangoproject_home: {your_djangoproject_home}")
# sys.path.append(your_djangoproject_home)
#
# from django.core.wsgi import get_wsgi_application
# os.environ.setdefault("DJANGO_SETTINGS_MODULE", "settings")
# application = get_wsgi_application()
# from sga.models import *
# from sagest.models import *
# from datetime import datetime
# import warnings
# warnings.filterwarnings('ignore', message='Unverified HTTPS request')
# print(u"Inicio")
# #solo posgrado graduados
# graduados = Graduado.objects.filter(status=True, inscripcion__carrera__coordinacion__id=7)
# try:
#     miarchivo = openpyxl.load_workbook("graduados_fecharefrendado_posgrado.xlsx")
#     lista = miarchivo.get_sheet_by_name('graduadosfecha')
#     totallista = lista.rows
#     a=0
#     c=0
#     listaerror = []
#     actualizados = 0
#     for filas in totallista:
#         try:
#             a += 1
#             if a > 1:
#                 id = str(filas[0].value).strip() if filas[0].value else None
#                 graduado = graduados.get(pk=id)
#                 if graduado.pk == 12457:
#                     print('ok')
#                 if graduado:
#                     fechacelda = str(filas[30].value).strip()
#                     if len(fechacelda)>0:
#                         if len(fechacelda) > 10:
#                             fechacelda = fechacelda[:len(fechacelda)-9]
#                         if '/' in fechacelda:
#                             fecha = datetime.strptime(fechacelda, '%d/%m/%Y').date()
#                         else:
#                             if len(str(fechacelda).split('-').__getitem__(0)) == 4:
#                                 fecha = datetime.strptime(fechacelda, '%Y-%m-%d').date()
#                             else:
#                                 fecha = datetime.strptime(fechacelda, '%d-%m-%Y').date()
#                         if fecha:
#                             # graduado.fecharefrendacion = date(int(fecha[0:4]), int(fecha[5:7]), int(fecha[8:10]))
#                             graduado.fecharefrendacion = date(fecha.year, fecha.month, fecha.day)
#                             graduado.save()
#                             actualizados = actualizados + 1
#         except Exception as ex:
#             print(f'error graduado {a}:, id: {id}, ex:{ex}')
#             listaerror.append(f'id:{id} - error:{str(ex)}')
# except Exception as ex:
#     print('error: %s' % ex)
# print('lista error', str(listaerror), ' total error: ', str(len(listaerror)))
# print('total actualizados', str(actualizados))
#
# # notificacion al administrativo en sga
# if len(listaerror) == 0:
#     titulonotificacion = f"Actualización exitosa"
#     cuerponotificacion = u"Se generó correctamente el proceso. \nTotal actualizados: %s. Errores: %s" % (str(actualizados), str(listaerror))
# else:
#     titulonotificacion = f"Error en el proceso de actualización"
#     cuerponotificacion = u"No se actualizaron %s registros: \n%s, \nTotal correctos: %s." % (
#         str(len(listaerror)), str(listaerror), str(actualizados))
# notificacion = Notificacion(
#     titulo=titulonotificacion,
#     cuerpo=cuerponotificacion,
#     destinatario=Persona.objects.get(pk=21966),
#     url=f"/graduados",
#     content_type=None,
#     object_id=None,
#     prioridad=1,
#     app_label='SGA',
#     fecha_hora_visible=datetime.now() + timedelta(days=3))
# notificacion.save()
# # fin notificacion al administrativo sga
# print("Finaliza")