import os
import sys
import io
import xlsxwriter
import xlwt
import openpyxl
import xlwt
from xlwt import *
from django.http import HttpResponse
from xlwt import *

# SITE_ROOT = os.path.dirname(os.path.realpath(__file__))
YOUR_PATH = os.path.dirname(os.path.realpath(__file__))
# print(f"YOUR_PATH: {YOUR_PATH}")
SITE_ROOT = os.path.dirname(os.path.dirname(YOUR_PATH))
SITE_ROOT = os.path.join(SITE_ROOT, '')
# print(f"SITE_ROOT: {SITE_ROOT}")
your_djangoproject_home = os.path.split(SITE_ROOT)[0]
# print(f"your_djangoproject_home: {your_djangoproject_home}")
sys.path.append(your_djangoproject_home)

from django.core.wsgi import get_wsgi_application

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "settings")
application = get_wsgi_application()

from django.http import HttpResponse
from settings import MEDIA_ROOT, BASE_DIR
from xlwt import easyxf, XFStyle
from sga.adm_criteriosactividadesdocente import asistencia_tutoria
from inno.models import *
from sga.models import *
from sagest.models import *
from inno.funciones import *
import concurrent.futures
from balcon.models import *
from Moodle_Funciones import crearhtmlphpmoodle

def cerrar_materias_transversales_5():
    materias = Materia.objects.filter(status=True, nivel__periodo_id=177, cerrado=False,
                                      asignaturamalla__malla__carrera__coordinacion__in=[5], modeloevaluativo_id=27)
    for materia in materias:
        for asig in materia.asignados_a_esta_materia():
            asig.cerrado = True
            asig.save(actualiza=False)
            asig.actualiza_estado()
        for asig in materia.asignados_a_esta_materia():
            asig.cierre_materia_asignada()

        materia.cerrado = True
        materia.fechacierre = datetime.now().date()
        materia.save()

#cerrar_materias_transversales_5()

def actualizar_nivel_inscripcion_malla9():
    matriculas = Matricula.objects.filter(status=True, nivel__periodo_id=177, inscripcion__carrera_id__in=[150,146,151,143,138,142,170,157,111,158] )
    for matricula in matriculas:
        inscripcion = matricula.inscripcion
        print('ACTUALIZANDO- ', inscripcion.persona.cedula)
        inscripcion.actualizar_nivel()
        print('ACTUALIZADO')
    print('FIN')


def inscribir_ppp_derecho():
    try:
        libre_origen = '/reporte_derecho_v4_1.xls'
        fuentecabecera = easyxf(
            'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
        fuentenormal = easyxf(
            'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')

        output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
        # liborigen = xlrd.open_workbook(output_folder + libre_origen)
        libdestino = xlwt.Workbook()
        hojadestino = libdestino.add_sheet('Sheet1')
        fil = 0
        lin = 0
        columnas = [(u"CEDULA", 6000),
                    (u"APELLIDOS Y NOMBRES", 6000),
                    (u"PRACTICAS", 6000),
                    (u"VINCULACION", 6000)
                    ]
        for col_num in range(len(columnas)):
            hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
            hojadestino.col(col_num).width = columnas[col_num][1]
        fila = 1

        miarchivo = openpyxl.load_workbook("derecholineaf.xlsx")
        # miarchivo = openpyxl.load_workbook("prueba_2.xlsx")

        ws = miarchivo.get_sheet_by_name("primero")
        worksheet = ws
        carrera_id = 126
        # mallaantigua_id = 198
        # mallanueva_id = 492
        curso = CursoEscuelaComplementaria.objects.get(pk=2434)

        for row in worksheet.iter_rows(min_row=0):
            if lin >= 0:
                currentValues, cadena = [], ''
                for cell in row:
                    cadena += str(cell.value) + ' '
                    currentValues.append(str(cell.value))
                identificacion = currentValues[0]

                if not identificacion:
                    break

                matricula = Matricula.objects.filter(status=True, inscripcion__carrera__id=carrera_id,
                                                     inscripcion__persona__cedula=identificacion).first()
                inscripcion = matricula.inscripcion
                hojadestino.write(fila, 0, inscripcion.persona.identificacion(), fuentenormal)
                hojadestino.write(fila, 1, inscripcion.persona.nombre_completo(), fuentenormal)
                practicaspp = haber_cumplido_horas_creditos_practicas_preprofesionales(inscripcion.id)
                horasvinculacion = haber_cumplido_horas_creditos_vinculacion(inscripcion.id)

                if horasvinculacion:
                    if not MatriculaCursoEscuelaComplementaria.objects.filter(curso=curso,
                                                                              inscripcion=inscripcion).exists():
                        registro = MatriculaCursoEscuelaComplementaria(curso=curso, inscripcion=inscripcion)
                        registro.save()
                        for materia in curso.materiacursoescuelacomplementaria_set.all():
                            asignatura = MateriaAsignadaCurso(inscripcion=registro,
                                                              materia=materia,
                                                              calificacion=100,
                                                              asistencia=100,
                                                              estado_id=3)
                            asignatura.save()
                    hojadestino.write(fila, 2, 'SI', fuentenormal)
                else:
                    hojadestino.write(fila, 2, 'NO', fuentenormal)
                if horasvinculacion:
                    hojadestino.write(fila, 3, 'SI', fuentenormal)
                else:
                    hojadestino.write(fila, 3, 'NO', fuentenormal)

                fila += 1

            lin += 1
        libdestino.save(output_folder + libre_origen)
        print("Proceso finalizado. . .")
    except Exception as ex:
        noti = Notificacion(titulo='Error',
                            cuerpo='Ha ocurrido un error {} - Error en la linea {}'.format(ex,
                                                                                           sys.exc_info()[
                                                                                               -1].tb_lineno),
                            destinatario_id=29898, url="",
                            prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                            tipo=2, en_proceso=False, error=True)
        noti.save()
        transaction.set_rollback(True)
        print('error: %s' % ex)
        hojadestino.write(fila, 4, str(ex))
        fila += 1

def inscribir_ppp_derecho_2():
    try:
        libre_origen = '/reporte_derecho_v4_2.xls'
        fuentecabecera = easyxf(
            'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
        fuentenormal = easyxf(
            'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')

        output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
        # liborigen = xlrd.open_workbook(output_folder + libre_origen)
        libdestino = xlwt.Workbook()
        hojadestino = libdestino.add_sheet('Sheet1')
        fil = 0
        lin = 0
        columnas = [(u"CEDULA", 6000),
                    (u"APELLIDOS Y NOMBRES", 6000),
                    (u"PRACTICAS", 6000),
                    (u"VINCULACION", 6000)
                    ]
        for col_num in range(len(columnas)):
            hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
            hojadestino.col(col_num).width = columnas[col_num][1]
        fila = 1

        miarchivo = openpyxl.load_workbook("derecholineaf.xlsx")
        # miarchivo = openpyxl.load_workbook("prueba_2.xlsx")

        ws = miarchivo.get_sheet_by_name("segundo")
        worksheet = ws
        carrera_id = 126
        # mallaantigua_id = 198
        # mallanueva_id = 492
        curso = CursoEscuelaComplementaria.objects.get(pk=2434)

        for row in worksheet.iter_rows(min_row=0):
            if lin >= 0:
                currentValues, cadena = [], ''
                for cell in row:
                    cadena += str(cell.value) + ' '
                    currentValues.append(str(cell.value))
                identificacion = currentValues[0]

                if not identificacion:
                    break

                matricula = Matricula.objects.filter(status=True, inscripcion__carrera__id=carrera_id,
                                                     inscripcion__persona__cedula=identificacion).first()
                inscripcion = matricula.inscripcion
                hojadestino.write(fila, 0, inscripcion.persona.identificacion(), fuentenormal)
                hojadestino.write(fila, 1, inscripcion.persona.nombre_completo(), fuentenormal)
                practicaspp = haber_cumplido_horas_creditos_practicas_preprofesionales(inscripcion.id)
                horasvinculacion = haber_cumplido_horas_creditos_vinculacion(inscripcion.id)

                if horasvinculacion:
                    if not MatriculaCursoEscuelaComplementaria.objects.filter(curso=curso,
                                                                              inscripcion=inscripcion).exists():
                        registro = MatriculaCursoEscuelaComplementaria(curso=curso, inscripcion=inscripcion)
                        registro.save()
                        for materia in curso.materiacursoescuelacomplementaria_set.all():
                            asignatura = MateriaAsignadaCurso(inscripcion=registro,
                                                              materia=materia,
                                                              calificacion=100,
                                                              asistencia=100,
                                                              estado_id=3)
                            asignatura.save()
                    hojadestino.write(fila, 2, 'SI', fuentenormal)
                else:
                    hojadestino.write(fila, 2, 'NO', fuentenormal)
                if horasvinculacion:
                    hojadestino.write(fila, 3, 'SI', fuentenormal)
                else:
                    hojadestino.write(fila, 3, 'NO', fuentenormal)

                fila += 1

            lin += 1
        libdestino.save(output_folder + libre_origen)
        print("Proceso finalizado. . .")
    except Exception as ex:
        noti = Notificacion(titulo='Error',
                            cuerpo='Ha ocurrido un error {} - Error en la linea {}'.format(ex,
                                                                                           sys.exc_info()[
                                                                                               -1].tb_lineno),
                            destinatario_id=29898, url="",
                            prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                            tipo=2, en_proceso=False, error=True)
        noti.save()
        transaction.set_rollback(True)
        print('error: %s' % ex)
        hojadestino.write(fila, 4, str(ex))
        fila += 1

def inscribir_ppp_derecho_3():
    try:
        libre_origen = '/reporte_derecho_v4_3.xls'
        fuentecabecera = easyxf(
            'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
        fuentenormal = easyxf(
            'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')

        output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
        # liborigen = xlrd.open_workbook(output_folder + libre_origen)
        libdestino = xlwt.Workbook()
        hojadestino = libdestino.add_sheet('Sheet1')
        fil = 0
        lin = 0
        columnas = [(u"CEDULA", 6000),
                    (u"APELLIDOS Y NOMBRES", 6000),
                    (u"PRACTICAS", 6000),
                    (u"VINCULACION", 6000)
                    ]
        for col_num in range(len(columnas)):
            hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
            hojadestino.col(col_num).width = columnas[col_num][1]
        fila = 1

        miarchivo = openpyxl.load_workbook("derecholineaf.xlsx")
        # miarchivo = openpyxl.load_workbook("prueba_2.xlsx")

        ws = miarchivo.get_sheet_by_name("tercero")
        worksheet = ws
        carrera_id = 126
        # mallaantigua_id = 198
        # mallanueva_id = 492
        curso = CursoEscuelaComplementaria.objects.get(pk=2434)

        for row in worksheet.iter_rows(min_row=0):
            if lin >= 0:
                currentValues, cadena = [], ''
                for cell in row:
                    cadena += str(cell.value) + ' '
                    currentValues.append(str(cell.value))
                identificacion = currentValues[0]

                if not identificacion:
                    break

                matricula = Matricula.objects.filter(status=True, inscripcion__carrera__id=carrera_id,
                                                     inscripcion__persona__cedula=identificacion).first()
                inscripcion = matricula.inscripcion
                hojadestino.write(fila, 0, inscripcion.persona.identificacion(), fuentenormal)
                hojadestino.write(fila, 1, inscripcion.persona.nombre_completo(), fuentenormal)
                practicaspp = haber_cumplido_horas_creditos_practicas_preprofesionales(inscripcion.id)
                horasvinculacion = haber_cumplido_horas_creditos_vinculacion(inscripcion.id)

                if horasvinculacion:
                    if not MatriculaCursoEscuelaComplementaria.objects.filter(curso=curso,
                                                                              inscripcion=inscripcion).exists():
                        registro = MatriculaCursoEscuelaComplementaria(curso=curso, inscripcion=inscripcion)
                        registro.save()
                        for materia in curso.materiacursoescuelacomplementaria_set.all():
                            asignatura = MateriaAsignadaCurso(inscripcion=registro,
                                                              materia=materia,
                                                              calificacion=100,
                                                              asistencia=100,
                                                              estado_id=3)
                            asignatura.save()
                    hojadestino.write(fila, 2, 'SI', fuentenormal)
                else:
                    hojadestino.write(fila, 2, 'NO', fuentenormal)
                if horasvinculacion:
                    hojadestino.write(fila, 3, 'SI', fuentenormal)
                else:
                    hojadestino.write(fila, 3, 'NO', fuentenormal)

                fila += 1

            lin += 1
        libdestino.save(output_folder + libre_origen)
        print("Proceso finalizado. . .")
    except Exception as ex:
        noti = Notificacion(titulo='Error',
                            cuerpo='Ha ocurrido un error {} - Error en la linea {}'.format(ex,
                                                                                           sys.exc_info()[
                                                                                               -1].tb_lineno),
                            destinatario_id=29898, url="",
                            prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                            tipo=2, en_proceso=False, error=True)
        noti.save()
        transaction.set_rollback(True)
        print('error: %s' % ex)
        hojadestino.write(fila, 4, str(ex))
        fila += 1

def inscribir_ppp_derecho_4():
    try:
        libre_origen = '/reporte_derecho_v4_4.xls'
        fuentecabecera = easyxf(
            'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
        fuentenormal = easyxf(
            'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')

        output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
        # liborigen = xlrd.open_workbook(output_folder + libre_origen)
        libdestino = xlwt.Workbook()
        hojadestino = libdestino.add_sheet('Sheet1')
        fil = 0
        lin = 0
        columnas = [(u"CEDULA", 6000),
                    (u"APELLIDOS Y NOMBRES", 6000),
                    (u"PRACTICAS", 6000),
                    (u"VINCULACION", 6000)
                    ]
        for col_num in range(len(columnas)):
            hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
            hojadestino.col(col_num).width = columnas[col_num][1]
        fila = 1

        miarchivo = openpyxl.load_workbook("derecholineaf.xlsx")
        # miarchivo = openpyxl.load_workbook("prueba_2.xlsx")

        ws = miarchivo.get_sheet_by_name("cuarto")
        worksheet = ws
        carrera_id = 126
        # mallaantigua_id = 198
        # mallanueva_id = 492
        curso = CursoEscuelaComplementaria.objects.get(pk=2434)

        for row in worksheet.iter_rows(min_row=0):
            if lin >= 0:
                currentValues, cadena = [], ''
                for cell in row:
                    cadena += str(cell.value) + ' '
                    currentValues.append(str(cell.value))
                identificacion = currentValues[0]

                if not identificacion:
                    break

                matricula = Matricula.objects.filter(status=True, inscripcion__carrera__id=carrera_id,
                                                     inscripcion__persona__cedula=identificacion).first()
                inscripcion = matricula.inscripcion
                hojadestino.write(fila, 0, inscripcion.persona.identificacion(), fuentenormal)
                hojadestino.write(fila, 1, inscripcion.persona.nombre_completo(), fuentenormal)
                practicaspp = haber_cumplido_horas_creditos_practicas_preprofesionales(inscripcion.id)
                horasvinculacion = haber_cumplido_horas_creditos_vinculacion(inscripcion.id)

                if horasvinculacion:
                    if not MatriculaCursoEscuelaComplementaria.objects.filter(curso=curso,
                                                                              inscripcion=inscripcion).exists():
                        registro = MatriculaCursoEscuelaComplementaria(curso=curso, inscripcion=inscripcion)
                        registro.save()
                        for materia in curso.materiacursoescuelacomplementaria_set.all():
                            asignatura = MateriaAsignadaCurso(inscripcion=registro,
                                                              materia=materia,
                                                              calificacion=100,
                                                              asistencia=100,
                                                              estado_id=3)
                            asignatura.save()
                    hojadestino.write(fila, 2, 'SI', fuentenormal)
                else:
                    hojadestino.write(fila, 2, 'NO', fuentenormal)
                if horasvinculacion:
                    hojadestino.write(fila, 3, 'SI', fuentenormal)
                else:
                    hojadestino.write(fila, 3, 'NO', fuentenormal)

                fila += 1

            lin += 1
        libdestino.save(output_folder + libre_origen)
        print("Proceso finalizado. . .")
    except Exception as ex:
        noti = Notificacion(titulo='Error',
                            cuerpo='Ha ocurrido un error {} - Error en la linea {}'.format(ex,
                                                                                           sys.exc_info()[
                                                                                               -1].tb_lineno),
                            destinatario_id=29898, url="",
                            prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                            tipo=2, en_proceso=False, error=True)
        noti.save()
        transaction.set_rollback(True)
        print('error: %s' % ex)
        hojadestino.write(fila, 4, str(ex))
        fila += 1

def inscribir_ppp_derecho_5():
    try:
        libre_origen = '/reporte_derecho_v4_5.xls'
        fuentecabecera = easyxf(
            'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
        fuentenormal = easyxf(
            'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')

        output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
        # liborigen = xlrd.open_workbook(output_folder + libre_origen)
        libdestino = xlwt.Workbook()
        hojadestino = libdestino.add_sheet('Sheet1')
        fil = 0
        lin = 0
        columnas = [(u"CEDULA", 6000),
                    (u"APELLIDOS Y NOMBRES", 6000),
                    (u"PRACTICAS", 6000),
                    (u"VINCULACION", 6000)
                    ]
        for col_num in range(len(columnas)):
            hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
            hojadestino.col(col_num).width = columnas[col_num][1]
        fila = 1

        miarchivo = openpyxl.load_workbook("derecholineaf.xlsx")
        # miarchivo = openpyxl.load_workbook("prueba_2.xlsx")

        ws = miarchivo.get_sheet_by_name("quinto")
        worksheet = ws
        carrera_id = 126
        # mallaantigua_id = 198
        # mallanueva_id = 492
        curso = CursoEscuelaComplementaria.objects.get(pk=2434)

        for row in worksheet.iter_rows(min_row=0):
            if lin >= 0:
                currentValues, cadena = [], ''
                for cell in row:
                    cadena += str(cell.value) + ' '
                    currentValues.append(str(cell.value))
                identificacion = currentValues[0]

                if not identificacion:
                    break

                matricula = Matricula.objects.filter(status=True, inscripcion__carrera__id=carrera_id,
                                                     inscripcion__persona__cedula=identificacion).first()
                inscripcion = matricula.inscripcion
                hojadestino.write(fila, 0, inscripcion.persona.identificacion(), fuentenormal)
                hojadestino.write(fila, 1, inscripcion.persona.nombre_completo(), fuentenormal)
                practicaspp = haber_cumplido_horas_creditos_practicas_preprofesionales(inscripcion.id)
                horasvinculacion = haber_cumplido_horas_creditos_vinculacion(inscripcion.id)

                if horasvinculacion:
                    if not MatriculaCursoEscuelaComplementaria.objects.filter(curso=curso,
                                                                              inscripcion=inscripcion).exists():
                        registro = MatriculaCursoEscuelaComplementaria(curso=curso, inscripcion=inscripcion)
                        registro.save()
                        for materia in curso.materiacursoescuelacomplementaria_set.all():
                            asignatura = MateriaAsignadaCurso(inscripcion=registro,
                                                              materia=materia,
                                                              calificacion=100,
                                                              asistencia=100,
                                                              estado_id=3)
                            asignatura.save()
                    hojadestino.write(fila, 2, 'SI', fuentenormal)
                else:
                    hojadestino.write(fila, 2, 'NO', fuentenormal)
                if horasvinculacion:
                    hojadestino.write(fila, 3, 'SI', fuentenormal)
                else:
                    hojadestino.write(fila, 3, 'NO', fuentenormal)

                fila += 1

            lin += 1
        libdestino.save(output_folder + libre_origen)
        print("Proceso finalizado. . .")
    except Exception as ex:
        noti = Notificacion(titulo='Error',
                            cuerpo='Ha ocurrido un error {} - Error en la linea {}'.format(ex,
                                                                                           sys.exc_info()[
                                                                                               -1].tb_lineno),
                            destinatario_id=29898, url="",
                            prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                            tipo=2, en_proceso=False, error=True)
        noti.save()
        transaction.set_rollback(True)
        print('error: %s' % ex)
        hojadestino.write(fila, 4, str(ex))
        fila += 1

def inscribir_ppp_derecho_6():
    try:
        libre_origen = '/reporte_derecho_v4_6.xls'
        fuentecabecera = easyxf(
            'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
        fuentenormal = easyxf(
            'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')

        output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
        # liborigen = xlrd.open_workbook(output_folder + libre_origen)
        libdestino = xlwt.Workbook()
        hojadestino = libdestino.add_sheet('Sheet1')
        fil = 0
        lin = 0
        columnas = [(u"CEDULA", 6000),
                    (u"APELLIDOS Y NOMBRES", 6000),
                    (u"PRACTICAS", 6000),
                    (u"VINCULACION", 6000)
                    ]
        for col_num in range(len(columnas)):
            hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
            hojadestino.col(col_num).width = columnas[col_num][1]
        fila = 1

        miarchivo = openpyxl.load_workbook("derecholineaf.xlsx")
        # miarchivo = openpyxl.load_workbook("prueba_2.xlsx")

        ws = miarchivo.get_sheet_by_name("sexto")
        worksheet = ws
        carrera_id = 126
        # mallaantigua_id = 198
        # mallanueva_id = 492
        curso = CursoEscuelaComplementaria.objects.get(pk=2434)

        for row in worksheet.iter_rows(min_row=0):
            if lin >= 0:
                currentValues, cadena = [], ''
                for cell in row:
                    cadena += str(cell.value) + ' '
                    currentValues.append(str(cell.value))
                identificacion = currentValues[0]

                if not identificacion:
                    break

                matricula = Matricula.objects.filter(status=True, inscripcion__carrera__id=carrera_id,
                                                     inscripcion__persona__cedula=identificacion).first()
                inscripcion = matricula.inscripcion
                hojadestino.write(fila, 0, inscripcion.persona.identificacion(), fuentenormal)
                hojadestino.write(fila, 1, inscripcion.persona.nombre_completo(), fuentenormal)
                practicaspp = haber_cumplido_horas_creditos_practicas_preprofesionales(inscripcion.id)
                horasvinculacion = haber_cumplido_horas_creditos_vinculacion(inscripcion.id)

                if horasvinculacion:
                    if not MatriculaCursoEscuelaComplementaria.objects.filter(curso=curso,
                                                                              inscripcion=inscripcion).exists():
                        registro = MatriculaCursoEscuelaComplementaria(curso=curso, inscripcion=inscripcion)
                        registro.save()
                        for materia in curso.materiacursoescuelacomplementaria_set.all():
                            asignatura = MateriaAsignadaCurso(inscripcion=registro,
                                                              materia=materia,
                                                              calificacion=100,
                                                              asistencia=100,
                                                              estado_id=3)
                            asignatura.save()
                    hojadestino.write(fila, 2, 'SI', fuentenormal)
                else:
                    hojadestino.write(fila, 2, 'NO', fuentenormal)
                if horasvinculacion:
                    hojadestino.write(fila, 3, 'SI', fuentenormal)
                else:
                    hojadestino.write(fila, 3, 'NO', fuentenormal)

                fila += 1

            lin += 1
        libdestino.save(output_folder + libre_origen)
        print("Proceso finalizado. . .")
    except Exception as ex:
        noti = Notificacion(titulo='Error',
                            cuerpo='Ha ocurrido un error {} - Error en la linea {}'.format(ex,
                                                                                           sys.exc_info()[
                                                                                               -1].tb_lineno),
                            destinatario_id=29898, url="",
                            prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                            tipo=2, en_proceso=False, error=True)
        noti.save()
        transaction.set_rollback(True)
        print('error: %s' % ex)
        hojadestino.write(fila, 4, str(ex))
        fila += 1

def inscribir_ppp_derecho_7():
    try:
        libre_origen = '/reporte_derecho_v4_7.xls'
        fuentecabecera = easyxf(
            'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
        fuentenormal = easyxf(
            'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')

        output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
        # liborigen = xlrd.open_workbook(output_folder + libre_origen)
        libdestino = xlwt.Workbook()
        hojadestino = libdestino.add_sheet('Sheet1')
        fil = 0
        lin = 0
        columnas = [(u"CEDULA", 6000),
                    (u"APELLIDOS Y NOMBRES", 6000),
                    (u"PRACTICAS", 6000),
                    (u"VINCULACION", 6000)
                    ]
        for col_num in range(len(columnas)):
            hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
            hojadestino.col(col_num).width = columnas[col_num][1]
        fila = 1

        miarchivo = openpyxl.load_workbook("derecholineaf.xlsx")
        # miarchivo = openpyxl.load_workbook("prueba_2.xlsx")

        ws = miarchivo.get_sheet_by_name("septimo")
        worksheet = ws
        carrera_id = 126
        # mallaantigua_id = 198
        # mallanueva_id = 492
        curso = CursoEscuelaComplementaria.objects.get(pk=2434)

        for row in worksheet.iter_rows(min_row=0):
            if lin >= 0:
                currentValues, cadena = [], ''
                for cell in row:
                    cadena += str(cell.value) + ' '
                    currentValues.append(str(cell.value))
                identificacion = currentValues[0]

                if not identificacion:
                    break

                matricula = Matricula.objects.filter(status=True, inscripcion__carrera__id=carrera_id,
                                                     inscripcion__persona__cedula=identificacion).first()
                inscripcion = matricula.inscripcion
                hojadestino.write(fila, 0, inscripcion.persona.identificacion(), fuentenormal)
                hojadestino.write(fila, 1, inscripcion.persona.nombre_completo(), fuentenormal)
                practicaspp = haber_cumplido_horas_creditos_practicas_preprofesionales(inscripcion.id)
                horasvinculacion = haber_cumplido_horas_creditos_vinculacion(inscripcion.id)

                if horasvinculacion:
                    if not MatriculaCursoEscuelaComplementaria.objects.filter(curso=curso,
                                                                              inscripcion=inscripcion).exists():
                        registro = MatriculaCursoEscuelaComplementaria(curso=curso, inscripcion=inscripcion)
                        registro.save()
                        for materia in curso.materiacursoescuelacomplementaria_set.all():
                            asignatura = MateriaAsignadaCurso(inscripcion=registro,
                                                              materia=materia,
                                                              calificacion=100,
                                                              asistencia=100,
                                                              estado_id=3)
                            asignatura.save()
                    hojadestino.write(fila, 2, 'SI', fuentenormal)
                else:
                    hojadestino.write(fila, 2, 'NO', fuentenormal)
                if horasvinculacion:
                    hojadestino.write(fila, 3, 'SI', fuentenormal)
                else:
                    hojadestino.write(fila, 3, 'NO', fuentenormal)

                fila += 1

            lin += 1
        libdestino.save(output_folder + libre_origen)
        print("Proceso finalizado. . .")
    except Exception as ex:
        noti = Notificacion(titulo='Error',
                            cuerpo='Ha ocurrido un error {} - Error en la linea {}'.format(ex,
                                                                                           sys.exc_info()[
                                                                                               -1].tb_lineno),
                            destinatario_id=29898, url="",
                            prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                            tipo=2, en_proceso=False, error=True)
        noti.save()
        transaction.set_rollback(True)
        print('error: %s' % ex)
        hojadestino.write(fila, 4, str(ex))
        fila += 1

def inscribir_ppp_derecho_8():
    try:
        libre_origen = '/reporte_derecho_v4_8.xls'
        fuentecabecera = easyxf(
            'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
        fuentenormal = easyxf(
            'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')

        output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
        # liborigen = xlrd.open_workbook(output_folder + libre_origen)
        libdestino = xlwt.Workbook()
        hojadestino = libdestino.add_sheet('Sheet1')
        fil = 0
        lin = 0
        columnas = [(u"CEDULA", 6000),
                    (u"APELLIDOS Y NOMBRES", 6000),
                    (u"PRACTICAS", 6000),
                    (u"VINCULACION", 6000)
                    ]
        for col_num in range(len(columnas)):
            hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
            hojadestino.col(col_num).width = columnas[col_num][1]
        fila = 1

        miarchivo = openpyxl.load_workbook("derecholineaf.xlsx")
        # miarchivo = openpyxl.load_workbook("prueba_2.xlsx")

        ws = miarchivo.get_sheet_by_name("octavo")
        worksheet = ws
        carrera_id = 126
        # mallaantigua_id = 198
        # mallanueva_id = 492
        curso = CursoEscuelaComplementaria.objects.get(pk=2434)

        for row in worksheet.iter_rows(min_row=0):
            if lin >= 0:
                currentValues, cadena = [], ''
                for cell in row:
                    cadena += str(cell.value) + ' '
                    currentValues.append(str(cell.value))
                identificacion = currentValues[0]

                if not identificacion:
                    break

                matricula = Matricula.objects.filter(status=True, inscripcion__carrera__id=carrera_id,
                                                     inscripcion__persona__cedula=identificacion).first()
                inscripcion = matricula.inscripcion
                hojadestino.write(fila, 0, inscripcion.persona.identificacion(), fuentenormal)
                hojadestino.write(fila, 1, inscripcion.persona.nombre_completo(), fuentenormal)
                practicaspp = haber_cumplido_horas_creditos_practicas_preprofesionales(inscripcion.id)
                horasvinculacion = haber_cumplido_horas_creditos_vinculacion(inscripcion.id)

                if horasvinculacion:
                    if not MatriculaCursoEscuelaComplementaria.objects.filter(curso=curso,
                                                                              inscripcion=inscripcion).exists():
                        registro = MatriculaCursoEscuelaComplementaria(curso=curso, inscripcion=inscripcion)
                        registro.save()
                        for materia in curso.materiacursoescuelacomplementaria_set.all():
                            asignatura = MateriaAsignadaCurso(inscripcion=registro,
                                                              materia=materia,
                                                              calificacion=100,
                                                              asistencia=100,
                                                              estado_id=3)
                            asignatura.save()
                    hojadestino.write(fila, 2, 'SI', fuentenormal)
                else:
                    hojadestino.write(fila, 2, 'NO', fuentenormal)
                if horasvinculacion:
                    hojadestino.write(fila, 3, 'SI', fuentenormal)
                else:
                    hojadestino.write(fila, 3, 'NO', fuentenormal)

                fila += 1

            lin += 1
        libdestino.save(output_folder + libre_origen)
        print("Proceso finalizado. . .")
    except Exception as ex:
        noti = Notificacion(titulo='Error',
                            cuerpo='Ha ocurrido un error {} - Error en la linea {}'.format(ex,
                                                                                           sys.exc_info()[
                                                                                               -1].tb_lineno),
                            destinatario_id=29898, url="",
                            prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                            tipo=2, en_proceso=False, error=True)
        noti.save()
        transaction.set_rollback(True)
        print('error: %s' % ex)
        hojadestino.write(fila, 4, str(ex))
        fila += 1

print("Función varios")
with concurrent.futures.ThreadPoolExecutor(max_workers=8) as executor:
    from settings import DEBUG
    # if DEBUG:
    #     path_anexo = 'reporte_acortezl__examen_admisión_enero2024.xlsx'
    future_1 = executor.submit(inscribir_ppp_derecho)
    future_2 = executor.submit(inscribir_ppp_derecho_2)
    future_3 = executor.submit(inscribir_ppp_derecho_3)
    future_4 = executor.submit(inscribir_ppp_derecho_4)
    future_5 = executor.submit(inscribir_ppp_derecho_5)
    future_6 = executor.submit(inscribir_ppp_derecho_6)
    future_7 = executor.submit(inscribir_ppp_derecho_7)
    future_8 = executor.submit(inscribir_ppp_derecho_8)