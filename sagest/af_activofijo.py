# -*- coding: UTF-8 -*-
import io
import json
import os
import sys
from datetime import datetime, timedelta
import openpyxl
import xlsxwriter
import xlwt
import pandas as pd
from xlwt import *
import random
import shutil
import zipfile
from unidecode import unidecode
from django.contrib import messages
from django.db.models import Sum
from googletrans import Translator
from django.contrib.auth.decorators import login_required
from django.db import transaction, connection
from django.db.models import Q, Max
from django.forms import model_to_dict
from django.http import HttpResponseRedirect, JsonResponse, HttpResponse
from django.shortcuts import render
from django.core.files import File as DjangoFile
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from django.template.loader import get_template
from openpyxl import load_workbook

from core.firmar_documentos_ec import JavaFirmaEc
from core.firmar_documentos_ec_descentralizada import qrImgFirma
from decorators import secure_module
from sagest.commonviews import secuencia_activos
from sagest.forms import ActivoFijoForm, InformeBajaForm, DetalleInformeBajaForm, \
    ImportarArchivoXLSForm, TraspasoActivoForm, BajaActivoForm, DetalleBajaActivoForm, AsignacionActivoForm, \
    TrasladoMantenimientoForm, MantenimientosActivosGarantiaForm, \
    DetalleNoIdentificadoForm, ActivosFijosForm, \
    DetalleTarjetaControlForm, ActasForm, ConsultacatalogoForm, ConsultausuarioForm, ConstatacionForm, \
    TraspasoActivoCustodioForm, ExportacionForm, EstadoBienForm, CondicionBienForm, EdificioForm, ArchivoActivoBajaForm, \
    ReporteEdificioFrom, TareaMantenimientoFrom, TipoBajaForm, TareaMantenimientoDaniosForm, \
    MantenimientosActivosPreventivosForm, GrupoBienForm, DirectorResponsableBajaForm, DirectorResponsableBaja2Form, \
    InformeBajaFormAF, PeriodoConstatacionAFForm, ConstatacionFisicaForm, DescargarBajasAniosForm, \
    GestionarBajaActivoForm, VerificacionTecnicaForm, ResponsableInformeBajaForm, DescargarActasConstatacionForm
from sagest.funciones import dominio_sistema_base, encrypt_id, formatear_cabecera_pd, generar_acta_constatacion, generar_acta_constatacion_reportlab
from sagest.models import ActivoFijo, ArchivoActivoFijo, HdDetalle_Incidente, MantenimientosActivosGarantia, \
    DetalleMantenimientosActivosGarantia, PersonaDepartamentoFirmas, \
    TraspasoActivo, DetalleTraspasoActivo, BajaActivo, DetalleBajaActivo, Ubicacion, \
    TrasladoMantenimiento, DetalleTrasladoMantenimiento, ConstatacionFisica, TIPO_MANTENIMIENTO, \
    DetalleConstatacionFisica, DetalleNoIdentificado, CuentaContable, DetalleMantenimiento, TareasActivosPreventivos, \
    CatalogoBien, ClaseVehiculo, TipoVehiculo, Color, EstadoProducto, OrigenIngreso, TipoDocumento, TarjetaControl, \
    ExportacionesActivos, Edificio, EstadoBien, CondicionBien, RangoVidaUtil, GruposCategoria, \
    MantenimientosActivosPreventivos, MantenimientoGruCategoria, TipoBaja, InformeActivoBaja, DetalleInformeActivoBaja, \
    MantenimientoGruCategoriaGarantiaLimp, MantenimientoGruCategoriaGarantiaErr, TareasActivosPreventivosGarantiaLimp, \
    TareasActivosPreventivosGarantiaErr, MantenimientoGruDanios, TareasActivosPreventivosDanios, \
    PiezaParteActivosPreventivos, HdPiezaPartes, ESTADO_DANIO, DirectorResponsableBaja, SolicitudActivos, \
    HistorialTraspaso, CLASE_BIEN, HistorialEstadoActivo, SolicitudTraspasoActivos, SeguimientoSolicitudTraspaso, \
    DistributivoPersona, PeriodoConstatacionAF, SeccionDepartamento, Departamento, DocumentoFirmaInformeBaja, \
    HistorialDocumentoInformeBaja, ESTADO_CONSTATACION, HistorialCertificadoFirmaPS, ActaConstatacion, \
    ESTADO_ACTA_CONSTATACION, ActivoTecnologico, GrupoDepartamento
from settings import RESPONSABLE_BIENES_ID, MEDIA_ROOT, ASISTENTE_BODEGA_ID, SITE_ROOT, SITE_STORAGE, MEDIA_URL
from sga.commonviews import adduserdata, obtener_reporte, traerNotificaciones
from sga.excelbackground import descarga_activos_bajas_background, reporte_activos_constatados_background, \
    reporte_activos_constatados_openxl_reportlab_background, reporte_historico_estado_activos_por_constatacion_background, \
    generar_acta_constatacion_background, descargar_informes_baja_masivo_background, descargar_actas_constatacion_masivo_background
from sga.templatetags.sga_extras import encrypt
from sga.funciones import MiPaginador, generar_nombre, log, convertir_fecha, convertir_fecha_invertida, \
    puede_realizar_accion, remover_caracteres_especiales_unicode
from sga.funciones import notificacion as notify
from sga.models import Persona, Reporte, Notificacion
from sga.reportes import elimina_tildes, run_report_v1
from sga.funcionesxhtml2pdf import conviert_html_to_pdf, conviert_html_to_pdfsaveqr_generico, \
    conviert_html_to_pdfsave_generic_lotes, conviert_html_to_pdf_save_file_model
from core.firmar_documentos import firmararchivogenerado, obtener_posicion_x_y_saltolinea
from utils.filtros_genericos import filtro_persona_select
from core.choices.models.sagest import MY_ESTADO_FIRMA_INFORME_BAJA
unicode = str


@login_required(redirect_field_name='ret', login_url='/loginsagest')
@secure_module
@transaction.atomic()
def view(request):
    data = {}
    adduserdata(request, data)
    usuario = request.user
    persona = request.session['persona']
    data['hoy']=hoy=datetime.now()
    data['DOMINIO_DEL_SISTEMA'] = dominio_sistema = dominio_sistema_base(request)

    if request.method == 'POST':
        action = request.POST['action']

        if action == 'add':
            try:
                f = ActivoFijoForm(request.POST)
                if f.is_valid():
                    if f.cleaned_data['responsable'] or f.cleaned_data['custodio'] or f.cleaned_data['ubicacion']:
                        if not f.cleaned_data['responsable'] or not f.cleaned_data['custodio'] or not f.cleaned_data[
                            'ubicacion']:
                            return JsonResponse(
                                {"result": "bad", "mensaje": u"Debe seleccionar Responsable, Custodio y Ubicación."})
                    activo = ActivoFijo(codigogobierno=f.cleaned_data['codigogobierno'],
                                        codigointerno=f.cleaned_data['codigointerno'],
                                        fechaingreso=f.cleaned_data['fechaingreso'],
                                        observacion=f.cleaned_data['observacion'],
                                        costo=f.cleaned_data['costo'],
                                        serie=f.cleaned_data['serie'],
                                        descripcion=f.cleaned_data['descripcion'],
                                        modelo=f.cleaned_data['modelo'],
                                        marca=f.cleaned_data['marca'],
                                        tipocomprobante=f.cleaned_data['tipocomprobante'],
                                        numerocomprobante=f.cleaned_data['numerocomprobante'],
                                        fechacomprobante=f.cleaned_data['fechacomprobante'],
                                        deprecia=f.cleaned_data['deprecia'],
                                        vidautil=f.cleaned_data['vidautil'],
                                        estructuraactivo=f.cleaned_data['estructuraactivo'],
                                        clasebien=f.cleaned_data['clasebien'],
                                        catalogo_id=f.cleaned_data['catalogo'],
                                        origeningreso=f.cleaned_data['origeningreso'],
                                        tipodocumentorespaldo=f.cleaned_data['tipodocumentorespaldo'],
                                        clasedocumentorespaldo=f.cleaned_data['clasedocumentorespaldo'],
                                        estado=f.cleaned_data['estado'],
                                        cuentacontable=f.cleaned_data['cuentacontable'],
                                        origenregistro=1,
                                        tipoproyecto=f.cleaned_data['tipoproyecto'])
                    activo.save(request)
                    if f.cleaned_data['responsable']:
                        activo.ubicacion = f.cleaned_data["ubicacion"]
                        activo.responsable_id = f.cleaned_data["responsable"]
                        activo.custodio_id = f.cleaned_data["custodio"]
                        traspaso = TraspasoActivo(usuariobienrecibe_id=f.cleaned_data['responsable'],
                                                  custodiobienrecibe_id=f.cleaned_data['custodio'],
                                                  ubicacionbienrecibe=f.cleaned_data['ubicacion'],
                                                  responsablebienes_id=RESPONSABLE_BIENES_ID,
                                                  fecha=datetime.now(),
                                                  tipo=1)
                        traspaso.save(request)
                        log(u'Adicionó traspaso: %s' % traspaso, request, "add")
                        secuencia = secuencia_activos(request)
                        secuencia.numeroasignacion += 1
                        secuencia.save(request)
                        traspaso.numero = secuencia.numeroasignacion
                        traspaso.save(request)
                        detalle = DetalleTraspasoActivo(codigotraspaso=traspaso,
                                                        activo=activo)
                        detalle.save(request)
                    if activo.catalogo.tipobien.id in [1, 3, 5]:
                        activo.color = f.cleaned_data['color']
                        activo.material = f.cleaned_data['material']
                        activo.dimensiones = f.cleaned_data['dimensiones']
                    elif activo.catalogo.tipobien.id == 2:
                        activo.colorprimario = f.cleaned_data['colorprimario']
                        activo.colorsecundario = f.cleaned_data['colorsecundario']
                        activo.clasevehiculo = f.cleaned_data['clasevehiculo']
                        activo.tipovehiculo = f.cleaned_data['tipovehiculo']
                        activo.numerochasis = f.cleaned_data['numerochasis']
                        activo.numeromotor = f.cleaned_data['numeromotor']
                        activo.placa = f.cleaned_data['placa']
                        activo.aniofabricacion = f.cleaned_data['aniofabricacion']
                    elif activo.catalogo.tipobien.id == 4:
                        activo.propietario = f.cleaned_data['propietario']
                        activo.codigocatastral = f.cleaned_data['codigocatastral']
                        activo.numeropredio = f.cleaned_data['numeropredio']
                        activo.valoravaluo = f.cleaned_data['valoravaluo']
                        activo.anioavaluo = f.cleaned_data['anioavaluo']
                        activo.areapredio = f.cleaned_data['areapredio']
                        activo.areaconstruccion = f.cleaned_data['areaconstruccion']
                        activo.pisos = f.cleaned_data['pisos']
                        activo.provincia = f.cleaned_data['provincia']
                        activo.canton = f.cleaned_data['canton']
                        activo.parroquia = f.cleaned_data['parroquia']
                        activo.zona = f.cleaned_data['zona']
                        activo.nomenclatura = f.cleaned_data['nomenclatura']
                        activo.sector = f.cleaned_data['sector']
                        activo.direccion = f.cleaned_data['direccion']
                        activo.escritura = f.cleaned_data['escritura']
                        activo.fechaescritura = f.cleaned_data['fechaescritura']
                        activo.notaria = f.cleaned_data['notaria']
                        activo.beneficiariocontrato = f.cleaned_data['beneficiariocontrato']
                        activo.fechacontrato = f.cleaned_data['fechacontrato']
                        activo.duracioncontrato = f.cleaned_data['duracioncontrato']
                        activo.montocontrato = f.cleaned_data['montocontrato']
                    else:
                        activo.titulo = f.cleaned_data['titulo']
                        activo.autor = f.cleaned_data['autor']
                        activo.editorial = f.cleaned_data['editorial']
                        activo.fechaedicion = f.cleaned_data['fechaedicion']
                        activo.numeroedicion = f.cleaned_data['numeroedicion']
                        activo.clasificacionbibliografica = f.cleaned_data['clasificacionbibliografica']
                    activo.save(request)
                    log(u'Adicionó Activo: %s' % activo, request, "add")
                    cuenta = CuentaContable.objects.get(pk=int(activo.cuentacontable_id))
                    cuenta.activosfijos = True
                    cuenta.save(request)
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'addingresoinformebaja':
            try:
                tipo = int(request.POST['tipo'])
                if InformeActivoBaja.objects.filter(tipoinforme=tipo, status=True,
                                                    activofijo_id=int(encrypt(request.POST['id']))).exists():
                    return JsonResponse({"result": "bad", "mensaje": u"El informe ya existe."})

                if tipo == 1:
                    f = InformeBajaForm(request.POST)

                else:
                    f = InformeBajaFormAF(request.POST)

                if f.is_valid():
                    if not f.cleaned_data['solicita']:
                        return JsonResponse({"result": "bad", "mensaje": u"Debe seleccionar solicitante."})
                    informebaja = InformeActivoBaja(activofijo_id=int(encrypt(request.POST['id'])),
                                                    solicita_id=f.cleaned_data['solicita'],
                                                    responsable_id=f.cleaned_data['responsable'],
                                                    conclusion=f.cleaned_data['conclusion'],
                                                    estado=f.cleaned_data['estado'],
                                                    estadouso=f.cleaned_data['estadouso'],
                                                    bloque=f.cleaned_data['bloque'],
                                                    detallerevision=f.cleaned_data['detallerevision'],
                                                    tipoinforme=tipo)
                    informebaja.save(request)

                    if tipo == 2:
                        informebaja.departamento = f.cleaned_data['departamento']
                        informebaja.gestion = f.cleaned_data['gestion']
                        informebaja.save(request)

                    if 'lista_items4' in request.POST:
                        listadetalle = json.loads(request.POST['lista_items4'])
                        if listadetalle:
                            for lisdet in listadetalle:
                                ingresodetalle = DetalleInformeActivoBaja(informebaja=informebaja,
                                                                          detalle=lisdet['listadetalle'])
                                ingresodetalle.save(request)

                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'edit':
            try:
                activo = ActivoFijo.objects.get(pk=request.POST['id'])
                f = ActivoFijoForm(request.POST)
                if f.is_valid():
                    activo.observacion = f.cleaned_data['observacion']
                    activo.serie = f.cleaned_data['serie']
                    activo.descripcion = f.cleaned_data['descripcion']
                    activo.modelo = f.cleaned_data['modelo']
                    activo.marca = f.cleaned_data['marca']
                    activo.tipocomprobante = f.cleaned_data['tipocomprobante']
                    activo.numerocomprobante = f.cleaned_data['numerocomprobante']
                    activo.clasebien = f.cleaned_data['clasebien']
                    activo.origeningreso = f.cleaned_data['origeningreso']
                    activo.tipodocumentorespaldo = f.cleaned_data['tipodocumentorespaldo']
                    activo.clasedocumentorespaldo = f.cleaned_data['clasedocumentorespaldo']
                    activo.estado = f.cleaned_data['estado']
                    activo.cuentacontable = f.cleaned_data['cuentacontable']
                    activo.tipoproyecto = f.cleaned_data['tipoproyecto']
                    activo.ubicacion = f.cleaned_data['ubicacion']
                    activo.fechainiciogarantia = f.cleaned_data['fechainiciogarantia']
                    activo.fechafingarantia = f.cleaned_data['fechafingarantia']
                    activo.ebye = f.cleaned_data['ebye']
                    if activo.custodio == None:
                        activo.custodio = Persona.objects.get(pk=f.cleaned_data['custodio'], status=True)
                    if activo.responsable == None:
                        activo.responsable = Persona.objects.get(pk=f.cleaned_data['responsable'], status=True)
                    if not activo.codigogobierno:
                        activo.codigogobierno = f.cleaned_data['codigogobierno']
                    if not activo.en_uso() and not activo.subidogobierno:
                        activo.fechaingreso = f.cleaned_data['fechaingreso']
                        activo.estructuraactivo = f.cleaned_data['estructuraactivo']
                        activo.costo = f.cleaned_data['costo']
                        activo.fechacomprobante = f.cleaned_data['fechacomprobante']
                    if not activo.en_uso() or request.user.has_perm("sagest.puede_modificar_depreciacion"):
                        activo.deprecia = f.cleaned_data['deprecia']
                        activo.vidautil = f.cleaned_data['vidautil']
                        catalago = activo.catalogo.id
                        if f.cleaned_data['catalogo']:
                            catalogo = CatalogoBien.objects.get(id=int(f.cleaned_data['catalogo']))
                        else:
                            catalogo = CatalogoBien.objects.get(id=activo.catalogo.id)
                        activo.catalogo = catalogo
                    activo.save(request)
                    if activo.catalogo.tipobien.id in [1, 3, 5]:
                        activo.color = f.cleaned_data['color']
                        activo.material = f.cleaned_data['material']
                        activo.dimensiones = f.cleaned_data['dimensiones']
                    elif activo.catalogo.tipobien.id == 2:
                        activo.colorprimario = f.cleaned_data['colorprimario']
                        activo.colorsecundario = f.cleaned_data['colorsecundario']
                        activo.clasevehiculo = f.cleaned_data['clasevehiculo']
                        activo.tipovehiculo = f.cleaned_data['tipovehiculo']
                        activo.numerochasis = f.cleaned_data['numerochasis']
                        activo.numeromotor = f.cleaned_data['numeromotor']
                        activo.placa = f.cleaned_data['placa']
                        activo.aniofabricacion = f.cleaned_data['aniofabricacion']
                    elif activo.catalogo.tipobien.id == 4:
                        activo.propietario = f.cleaned_data['propietario']
                        activo.codigocatastral = f.cleaned_data['codigocatastral']
                        activo.numeropredio = f.cleaned_data['numeropredio']
                        activo.valoravaluo = f.cleaned_data['valoravaluo']
                        activo.anioavaluo = f.cleaned_data['anioavaluo']
                        activo.areapredio = f.cleaned_data['areapredio']
                        activo.areaconstruccion = f.cleaned_data['areaconstruccion']
                        activo.pisos = f.cleaned_data['pisos']
                        activo.provincia = f.cleaned_data['provincia']
                        activo.canton = f.cleaned_data['canton']
                        activo.parroquia = f.cleaned_data['parroquia']
                        activo.zona = f.cleaned_data['zona']
                        activo.nomenclatura = f.cleaned_data['nomenclatura']
                        activo.sector = f.cleaned_data['sector']
                        activo.direccion = f.cleaned_data['direccion']
                        activo.escritura = f.cleaned_data['escritura']
                        activo.fechaescritura = f.cleaned_data['fechaescritura']
                        activo.notaria = f.cleaned_data['notaria']
                        activo.beneficiariocontrato = f.cleaned_data['beneficiariocontrato']
                        activo.fechacontrato = f.cleaned_data['fechacontrato']
                        activo.duracioncontrato = f.cleaned_data['duracioncontrato']
                        activo.montocontrato = f.cleaned_data['montocontrato']
                    else:
                        activo.titulo = f.cleaned_data['titulo']
                        activo.autor = f.cleaned_data['autor']
                        activo.editorial = f.cleaned_data['editorial']
                        activo.fechaedicion = f.cleaned_data['fechaedicion']
                        activo.numeroedicion = f.cleaned_data['numeroedicion']
                        activo.clasificacionbibliografica = f.cleaned_data['clasificacionbibliografica']
                    activo.save(request)
                    log(u'Editó Activo: %s' % activo, request, "edit")
                    fechaultimadeprec = ActivoFijo.objects.aggregate(fecha=Max('fechaultimadeprec'))['fecha']
                    activo.depreciar_activo(fechaultimadeprec, request)
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'gestionarbaja':
            try:
                f = GestionarBajaActivoForm(request.POST)

                if not f.is_valid():
                    for k, v in f.errors.items():
                        raise NameError(k + ', ' + v[0])

                # Consulto activo fijo
                activo = ActivoFijo.objects.get(pk=int(encrypt(request.POST['id'])))

                ubicacionbodega = f.cleaned_data['ubicacionbodega']
                historial = HistorialEstadoActivo(activo=activo, estado=activo.estado,
                                                  condicionestado=activo.condicionestado,
                                                  persona=persona, enuso=activo.enuso,
                                                  tiporegistro=3, observacion='GESTIÓN DE BAJA',
                                                  gestionbaja=True, ubicacionbodega=ubicacionbodega)
                historial.save(request)
                activo.procesobaja = esta = True
                activo.ubicacionbodega = ubicacionbodega
                # activo.estado = f.cleaned_data['estado']
                activo.save(request)

                periodoc = PeriodoConstatacionAF.objects.filter(status=True, activo=True).first()

                if periodoc:
                    detalle_c = DetalleConstatacionFisica.objects.filter(status=True, activo=activo,
                                                                         codigoconstatacion__periodo=periodoc).first()
                    if not detalle_c:
                        constatacion = ConstatacionFisica.objects.filter(status=True, usuariobienes=activo.responsable,
                                                                         periodo=periodoc).first()
                        if not constatacion:
                            secuencia = secuencia_activos(request)
                            secuencia.numeroconstatacion += 1
                            secuencia.save(request)
                            constatacion = ConstatacionFisica(usuariobienes=activo.responsable,
                                                              numero=secuencia.numeroconstatacion,
                                                              normativaconstatacion=secuencia.normativaconstatacion,
                                                              fechainicio=hoy,
                                                              periodo=periodoc,
                                                              ubicacionbienes=activo.ubicacion)
                            constatacion.save(request)
                            log(u'Agrego constatación: %s' % (constatacion), request, "add")
                        detalle_c = DetalleConstatacionFisica(codigoconstatacion=constatacion,
                                                              activo=activo,
                                                              responsable=persona,
                                                              estadooriginal=activo.estado)
                        detalle_c.save(request)
                        log(u'Agrego detalle de constatación: %s' % (detalle_c), request, "add")

                    detalle_c.encontrado = True
                    detalle_c.enuso = False
                    detalle_c.estadoactual = activo.estado
                    detalle_c.ubicacionbienes = activo.ubicacion
                    detalle_c.usuariobienes = activo.responsable
                    detalle_c.requiredarbaja = True
                    detalle_c.save(request)

                log(u'Inicia proceso de baja : %s (%s)' % (activo, activo.procesobaja), request, "edit")
                return JsonResponse({"result": False}, safe=False)
            except Exception as ex:
                msg = ex.__str__()
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos. [%s]" % (msg)})

        elif action == 'quitargestionarbaja':
            try:
                activo = ActivoFijo.objects.get(pk=encrypt_id(request.POST['id']))
                activo.procesobaja = False
                activo.ubicacionbodega = ''
                activo.save(request)
                historial = HistorialEstadoActivo(activo=activo, estado=activo.estado,
                                                  condicionestado=activo.condicionestado,
                                                  persona=persona, enuso=activo.enuso,
                                                  tiporegistro=3, observacion='CANCELA GESTIÓN DE BAJA',
                                                  gestionbaja=False)

                historial.save(request)
                log(u'Cancela proceso de baja : %s (%s)' % (activo, activo.procesobaja), request, "edit")
                return JsonResponse({"result": 'ok'})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'editingresoinformebaja':
            try:
                informebaja = InformeActivoBaja.objects.get(pk=int(request.POST['id']))
                f = InformeBajaFormAF(request.POST)
                if informebaja.tipoinforme == 1:
                    f = InformeBajaForm(request.POST)
                if f.is_valid():
                    if not f.cleaned_data['solicita']:
                        return JsonResponse({"result": "bad", "mensaje": u"Debe seleccionar solicitante."})
                    informebaja.solicita_id = f.cleaned_data['solicita']
                    informebaja.responsable_id = f.cleaned_data['responsable']
                    informebaja.detallerevision = f.cleaned_data['detallerevision']
                    informebaja.bloque = f.cleaned_data['bloque']
                    informebaja.estadouso = f.cleaned_data['estadouso']
                    informebaja.estado = f.cleaned_data['estado']
                    informebaja.conclusion = f.cleaned_data['conclusion']
                    if informebaja.tipoinforme == 2:
                        informebaja.departamento = f.cleaned_data['departamento']
                        informebaja.gestion = f.cleaned_data['gestion']
                    informebaja.save(request)

                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'eliminarinformebaja':
            try:
                with transaction.atomic():
                    informebaja = InformeActivoBaja.objects.get(pk=int(request.POST['id']))
                    informebaja.status = False
                    informebaja.save(request)

                    for detalleinformebaja in DetalleInformeActivoBaja.objects.filter(informebaja = informebaja, status = True):
                        detalleinformebaja.status = False
                        detalleinformebaja.save(request)

                    log(u'Eliminacion de informe de baja: %s' % informebaja, request, "eliminarinformebaja")
                    res_json = {"error": False}
            except Exception as ex:
                transaction.set_rollback(True)
                res_json = {'error': True, "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)

        elif action == 'additemdetalleinformebaja':
            try:
                detalleinf = DetalleInformeActivoBaja(informebaja_id=request.POST['idactivoinformebaja'],
                                                      detalle=request.POST['descripcion'])
                detalleinf.save(request)
                log(u'Adicinó detalle de informe de baja: %s' % detalleinf, request, "add")
                return JsonResponse({"result": "ok", 'codigoinformedet': detalleinf.id})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'listainformebaja':
            try:
                detalleinf = DetalleInformeActivoBaja.objects.get(pk=request.POST['id'])
                descripcion = detalleinf.detalle
                idobjetivo = detalleinf.id
                return JsonResponse({"result": "ok", 'descripcion': descripcion, 'codigorai': idobjetivo})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'edititemdetalleinformebaja':
            try:
                detalleinf = DetalleInformeActivoBaja.objects.get(pk=request.POST['codigoitemdetalleinformebaja'])
                detalleinf.detalle = request.POST['descripcion']
                detalleinf.save(request)
                log(u'Modifico detalle informe de baja: %s ' % detalleinf, request, "edit")
                return JsonResponse({"result": "ok", "descripcion": detalleinf.detalle})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'itemdetalleinformebaja':
            try:
                detalleinf = DetalleInformeActivoBaja.objects.get(pk=request.POST['id'])
                descripcion = detalleinf.detalle
                idobjetivo = detalleinf.id
                return JsonResponse({"result": "ok", 'descripcion': descripcion, 'codigorai': idobjetivo})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'eliminaritemdetalleinformebaja':
            try:
                itemdetalle = DetalleInformeActivoBaja.objects.get(pk=request.POST['idcodigodet'])
                log(u'Eliminó item detalle informe baja: %s' % itemdetalle, request, "del")
                itemdetalle.delete()
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        elif action == 'editmantenimientogarantia':
            try:
                mangarantia = MantenimientosActivosGarantia.objects.get(pk=int(request.POST['id']))
                limp = request.POST['id_limpiar']
                danio = request.POST['id_danio']
                if 'archivo' in request.FILES:
                    d = request.FILES['archivo']
                    if d.size > 2194304:
                        return JsonResponse({"result": "bad", "mensaje": u"Error, archivo mayor a 2 Mb."})
                    else:
                        newfiles = request.FILES['archivo']
                        newfilesd = newfiles._name
                        ext = newfilesd[newfilesd.rfind("."):]
                        if not (ext == '.pdf'):
                            return JsonResponse({"result": "bad", "mensaje": u"Error, solo archivos .pdf"})
                if 'arcusen' in request.FILES:
                    d = request.FILES['arcusen']
                    if d.size > 2194304:
                        return JsonResponse({"result": "bad", "mensaje": u"Error, archivo mayor a 2 Mb."})
                    else:
                        newfiles = request.FILES['arcusen']
                        newfilesd = newfiles._name
                        ext = newfilesd[newfilesd.rfind("."):]
                        if not (ext == '.pdf'):
                            return JsonResponse({"result": "bad", "mensaje": u"Error, solo archivos .pdf"})
                f = MantenimientosActivosGarantiaForm(request.POST, request.FILES)
                if f.is_valid():
                    # mangarantia.activofijo_id = f.cleaned_data['activofijo']
                    mangarantia.proveedor_id = f.cleaned_data['proveedor']
                    mangarantia.valor = f.cleaned_data['valor']
                    mangarantia.numreporte = f.cleaned_data['numreporte']
                    mangarantia.fechainicio = f.cleaned_data['fechainicio']
                    mangarantia.horamax = f.cleaned_data['horamax']
                    mangarantia.minutomax = f.cleaned_data['minutomax']
                    mangarantia.estfrec = f.cleaned_data['estfrec']
                    mangarantia.estfent = f.cleaned_data['estfent']
                    mangarantia.observacion = f.cleaned_data['observacion']
                    mangarantia.estusu = f.cleaned_data['estusu']
                    mangarantia.save(request)
                    if 'archivo' in request.FILES:
                        newfile = request.FILES['archivo']
                        newfile._name = generar_nombre("archivo_garantia", newfile._name)
                        mangarantia.archivo = newfile
                        # capacitacion.tiempo = f.cleaned_data['tiempo']
                        mangarantia.save(request)
                    if 'arcusen' in request.FILES:
                        newfile = request.FILES['arcusen']
                        newfile._name = generar_nombre("archivo_entusuario", newfile._name)
                        mangarantia.arcusen = newfile
                        # capacitacion.tiempo = f.cleaned_data['tiempo']
                        mangarantia.save(request)

                    TareasActivosPreventivosGarantiaLimp.objects.filter(mantenimiento=mangarantia).delete()
                    elementoslimp = limp.split(',')
                    for elemento in elementoslimp:
                        detalle = TareasActivosPreventivosGarantiaLimp(mantenimiento=mangarantia, grupos_id=elemento)
                        detalle.save(request)

                    TareasActivosPreventivosGarantiaErr.objects.filter(mantenimiento=mangarantia).delete()
                    elementoserr = danio.split(',')
                    for elemento in elementoserr:
                        detalle = TareasActivosPreventivosGarantiaErr(mantenimiento=mangarantia, grupos_id=elemento)
                        detalle.save(request)
                    log(u'Modifico mantenimiento garantia: %s' % mangarantia, request, "edit")
                    return JsonResponse({'result': 'ok'})
                else:
                    return JsonResponse({"result": "bad", "mensaje": u"Error, al guardar los datos."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', 'mensaje': u'Error al guardar los datos'})

        elif action == 'asignar':
            try:
                f = AsignacionActivoForm(request.POST)
                if f.is_valid():
                    items = json.loads(request.POST['lista_items1'])
                    if len(items) == 0:
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"Debe seleccionar activos para poder asignar"})
                    traspaso = TraspasoActivo(usuariobienrecibe=f.cleaned_data['usuariobienrecibe'],
                                              custodiobienrecibe=f.cleaned_data['custodiobienrecibe'],
                                              ubicacionbienrecibe=f.cleaned_data['ubicacionbienrecibe'],
                                              fecha=datetime.now(),
                                              tipo=1)
                    traspaso.save(request)
                    secuencia = secuencia_activos(request)
                    secuencia.numeroasignacion += 1
                    secuencia.save(request)
                    traspaso.numero = secuencia.numeroasignacion
                    traspaso.save(request)
                    for d in items:
                        activo = ActivoFijo.objects.get(pk=int(d['id']))
                        detalle = DetalleTraspasoActivo(codigotraspaso=traspaso,
                                                        activo=activo)
                        detalle.save(request)
                        activo.actualiza_responsable()

                    log(u'Asigno Activo: %s' % traspaso, request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'listatipomantenimiento':
            try:
                listatareas = MantenimientoGruCategoria.objects.filter(grupocategoria_id=request.POST['id_tipo'],
                                                                       status=True, activo=True)
                lista = []
                for lis in listatareas:
                    lista.append([lis.id, lis.descripcion])
                return JsonResponse({'result': 'ok', 'lista': lista})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'listatipomantenimientosgdanio':
            try:
                listatareas = MantenimientoGruDanios.objects.filter(grupocategoria_id=request.POST['id_tipo'],
                                                                    status=True, activo=True)
                lista = []
                for lis in listatareas:
                    lista.append([lis.id, lis.descripcion])
                return JsonResponse({'result': 'ok', 'lista': lista})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'afmarcaymodelo':
            try:
                listatareas = ActivoFijo.objects.filter(id=request.POST['id_tipo'], status=True)
                lista = []
                for lis in listatareas:
                    lista.append([lis.marca, lis.modelo])
                return JsonResponse({'result': 'ok', 'lista': lista})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'addmantenimiento':
            try:
                form = MantenimientosActivosPreventivosForm(request.POST, request.FILES)
                if 'archivo' in request.FILES:
                    newfile = request.FILES['archivo']
                    if newfile:
                        newfilesd = newfile._name
                        ext = newfilesd[newfilesd.rfind("."):]
                        if not ext.lower() == '.pdf':
                            return JsonResponse({"result": "bad", "mensaje": u"Error, solo archivos .pdf."})
                        if newfile.size > 2194304:
                            return JsonResponse({"result": "bad", "mensaje": u"Error, archivo mayor a 2 Mb."})
                        if newfile:
                            newfile._name = generar_nombre("archivo_malla_", newfile._name)
                if form.is_valid():
                    id_cpu = request.POST['cpu']
                    id_tipact = request.POST['tipoactivo']
                    estusu = form.cleaned_data['estusu']
                    fecha = form.cleaned_data['fecha']
                    horamax = form.cleaned_data['horamax']
                    minutomax = form.cleaned_data['minutomax']
                    estfrec = form.cleaned_data['estfrec']
                    estfent = form.cleaned_data['estfent']
                    marca = form.cleaned_data['marca']
                    modelo = form.cleaned_data['modelo']
                    bsugiere = form.cleaned_data['bsugiere']
                    dsugiere = form.cleaned_data['dsugiere']
                    observacion = form.cleaned_data['observacion']
                    limp = request.POST['id_limp']
                    # danio = request.POST['id_dani']
                    piezaparte = request.POST.getlist('piezaparte[]')
                    caracteristica = request.POST.getlist('catacteristica[]')
                    danio = request.POST.getlist('danio[]')
                    daniodes = request.POST.getlist('danioop[]')

                    mantenimientosactivos = MantenimientosActivosPreventivos(activofijo_id=id_cpu,
                                                                             tipoactivo_id=id_tipact,
                                                                             estusu=estusu,
                                                                             fecha=fecha,
                                                                             horamax=horamax,
                                                                             minutomax=minutomax,
                                                                             funcionarecibe=estfrec,
                                                                             funcionaentrega=estfent,
                                                                             marca=marca,
                                                                             modelo=modelo,
                                                                             sbequipo=bsugiere,
                                                                             descbaja=dsugiere,
                                                                             observaciones=observacion,
                                                                             nuevo=True)
                    mantenimientosactivos.save(request)
                    limp = limp.split(',')
                    for elemento in limp:
                        detalle = TareasActivosPreventivos(mantenimiento=mantenimientosactivos,
                                                           grupos_id=elemento)
                        detalle.save(request)
                    c = 0
                    while c < len(danio):
                        detalle = TareasActivosPreventivosDanios(mantenimiento=mantenimientosactivos,
                                                                 grupos_id=int(MantenimientoGruDanios.objects.get(
                                                                     descripcion=danio[c], grupocategoria_id=id_tipact,
                                                                     status=True).id),
                                                                 estadodanio=daniodes[c])
                        detalle.save(request)
                        c += 1
                    counter = 0
                    while counter < len(piezaparte):
                        detalle = PiezaParteActivosPreventivos(mantenimiento=mantenimientosactivos,
                                                               piezaparte_id=int(HdPiezaPartes.objects.get(
                                                                   descripcion=piezaparte[counter],
                                                                   grupocategoria_id=id_tipact, status=True).id),
                                                               descripcion=caracteristica[counter])
                        detalle.save(request)
                        counter += 1
                    if 'archivo' in request.FILES:
                        newfile = request.FILES['archivo']
                        if newfile:
                            newfile._name = generar_nombre("evidencia_comunicado", newfile._name)
                            mantenimientosactivos.archivo = newfile
                            mantenimientosactivos.save(request)
                    log(u'Asigno mantenimiento preventivo Activo: %s' % mantenimientosactivos, request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'iniciabaja':
            try:
                activo = ActivoFijo.objects.get(pk=request.POST['id'])

                if request.POST['val'] == 'y':
                    historial = HistorialEstadoActivo(activo=activo, estado=activo.estado)
                    historial.save(request)
                    activo.procesobaja = esta = True
                    activo.estado_id = 3
                    activo.save(request)

                    periodoc = PeriodoConstatacionAF.objects.filter(status=True,activo=True).first()

                    detalle_c = DetalleConstatacionFisica.objects.filter(status=True, activo=activo,
                                                                         codigoconstatacion__periodo=periodoc).first()
                    if not detalle_c:
                        constatacion = ConstatacionFisica.objects.filter(status=True, usuariobienes=activo.responsable,
                                                                         periodo=periodoc).first()
                        if not constatacion:
                            secuencia = secuencia_activos(request)
                            secuencia.numeroconstatacion += 1
                            secuencia.save(request)
                            constatacion = ConstatacionFisica(usuariobienes=activo.responsable,
                                                              numero=secuencia.numeroconstatacion,
                                                              normativaconstatacion=secuencia.normativaconstatacion,
                                                              fechainicio=hoy,
                                                              periodo=periodoc,
                                                              ubicacionbienes=activo.ubicacion)
                            constatacion.save(request)
                            log(u'Agrego constatación: %s' % (constatacion), request, "add")
                        detalle_c = DetalleConstatacionFisica(codigoconstatacion=constatacion,
                                                              activo=activo,
                                                              responsable=persona,
                                                              estadooriginal=activo.estado)
                        detalle_c.save(request)
                        log(u'Agrego detalle de constatación: %s' % (detalle_c), request, "add")

                    detalle_c.encontrado = True
                    detalle_c.enuso = False
                    detalle_c.estadoactual = activo.estado
                    detalle_c.ubicacionbienes = activo.ubicacion
                    detalle_c.usuariobienes = activo.responsable
                    detalle_c.requiredarbaja = True
                    detalle_c.save(request)

                else:
                    activo.procesobaja = esta = False
                    activo.ubicacionbodega = ''
                    # historial = HistorialEstadoActivo.objects.filter(activo=activo, status=True).order_by('-id').last()
                    # if historial:
                        # activo.estado_id = historial.estado
                        # DetalleConstatacionFisica.objects.filter(status=True, codigoconstatacion__estado=1,
                        #                                          activo=activo).update(estadoactual=historial.estado)

                activo.save(request)
                historial = HistorialEstadoActivo(activo=activo, estado=activo.estado,
                                                  condicionestado=activo.condicionestado,
                                                  persona=persona, enuso=activo.enuso,
                                                  tiporegistro=3, observacion='CANCELA GESTIÓN DE BAJA',
                                                  gestionbaja=False)

                historial.save(request)
                log(u'Edita proceso de baja : %s (%s)' % (activo, activo.procesobaja), request, "edit")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad"})

        elif action == 'editmantenimiento':
            try:
                f = AsignacionActivoForm(request.POST)
                if f.is_valid():
                    id_activoman = request.POST['id_activoman']
                    id_monitor = request.POST['id_monitor']
                    id_mouse = request.POST['id_mouse']
                    id_teclado = request.POST['id_teclado']
                    id_fechaingreso = request.POST['id_fechaingreso']
                    id_mantenimiento = request.POST['id_mantenimiento']
                    id_tipo = request.POST['id_tipo']
                    id_procesador = request.POST['id_procesador']
                    id_memoria = request.POST['id_memoria']
                    id_discoduro = request.POST['id_discoduro']
                    id_particiones = request.POST['id_particiones']
                    id_sistemaoperativo = request.POST['id_sistemaoperativo']
                    id_arquitectura = request.POST['id_arquitectura']
                    id_service = request.POST['id_service']
                    id_observacion = request.POST['id_observacion']
                    lista = request.POST['lista']
                    id_recibe = request.POST['id_recibe']
                    recibeid = False
                    if id_recibe == '1':
                        recibeid = True
                    id_entrega = request.POST['id_entrega']
                    entregaid = False
                    if id_entrega == '1':
                        entregaid = True
                    mantenimientosactivos = MantenimientosActivosPreventivos.objects.get(pk=id_activoman)
                    mantenimientosactivos.tipomantenimiento = id_mantenimiento
                    monitor = id_monitor
                    mantenimientosactivos.mouse = id_mouse
                    mantenimientosactivos.teclado = id_teclado
                    mantenimientosactivos.procesador = id_procesador
                    mantenimientosactivos.memoria = id_memoria
                    mantenimientosactivos.discoduro = id_discoduro
                    mantenimientosactivos.particiones = id_particiones
                    mantenimientosactivos.sistemaoperativo = id_sistemaoperativo
                    mantenimientosactivos.arquitectura = id_arquitectura
                    mantenimientosactivos.service = id_service
                    mantenimientosactivos.fecha = id_fechaingreso
                    mantenimientosactivos.tipoactivo_id = id_tipo
                    mantenimientosactivos.funcionarecibe = recibeid
                    mantenimientosactivos.funcionaentrega = entregaid
                    mantenimientosactivos.observaciones = id_observacion
                    mantenimientosactivos.save(request)

                    TareasActivosPreventivos.objects.filter(mantenimiento=mantenimientosactivos).delete()
                    elementos = lista.split(',')
                    for elemento in elementos:
                        detalle = TareasActivosPreventivos(mantenimiento=mantenimientosactivos,
                                                           grupos_id=elemento)
                        detalle.save(request)
                    log(u'Editó mantenimiento preventivo Activo: %s' % mantenimientosactivos, request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'editmantenimientov2':
            try:
                form = MantenimientosActivosPreventivosForm(request.POST, request.FILES)
                if form.is_valid():
                    eli_arc = request.POST['archivo-clear'] if 'archivo-clear' in request.POST else ''
                    id_activoman = request.POST['id']
                    id_cpu = request.POST['cpu']
                    id_tipact = request.POST['tipoactivo']
                    estusu = form.cleaned_data['estusu']
                    fecha = form.cleaned_data['fecha']
                    horamax = form.cleaned_data['horamax']
                    minutomax = form.cleaned_data['minutomax']
                    estfrec = form.cleaned_data['estfrec']
                    estfent = form.cleaned_data['estfent']
                    marca = form.cleaned_data['marca']
                    modelo = form.cleaned_data['modelo']
                    bsugiere = form.cleaned_data['bsugiere']
                    dsugiere = form.cleaned_data['dsugiere']
                    observacion = form.cleaned_data['observacion']
                    limp = request.POST['id_limpiar']
                    # danio = request.POST['id_danio']
                    piezaparte = request.POST.getlist('piezaparte[]')
                    caracteristica = request.POST.getlist('catacteristica[]')
                    danio = request.POST.getlist('danio[]')
                    daniodes = request.POST.getlist('danioop[]')

                    mantenimientosactivos = MantenimientosActivosPreventivos.objects.get(id=id_activoman)
                    mantenimientosactivos.activofijo_id = id_cpu
                    mantenimientosactivos.tipoactivo_id = id_tipact
                    mantenimientosactivos.estusu = estusu
                    mantenimientosactivos.fecha = fecha
                    mantenimientosactivos.horamax = horamax
                    if eli_arc == 'on':
                        mantenimientosactivos.archivo = None
                    mantenimientosactivos.minutomax = minutomax
                    mantenimientosactivos.funcionarecibe = estfrec
                    mantenimientosactivos.funcionaentrega = estfent
                    mantenimientosactivos.marca = marca
                    mantenimientosactivos.modelo = modelo
                    mantenimientosactivos.sbequipo = bsugiere
                    mantenimientosactivos.descbaja = dsugiere
                    mantenimientosactivos.observaciones = observacion
                    mantenimientosactivos.save(request)
                    if 'archivo' in request.FILES:
                        newfile = request.FILES['archivo']
                        newfile._name = generar_nombre("evidencia_comunicado", newfile._name)
                        mantenimientosactivos.archivo = newfile
                        mantenimientosactivos.save(request)
                    TareasActivosPreventivos.objects.filter(mantenimiento=mantenimientosactivos).delete()
                    elementos = limp.split(',')
                    for elemento in elementos:
                        detalle = TareasActivosPreventivos(mantenimiento=mantenimientosactivos,
                                                           grupos_id=elemento)
                        detalle.save(request)
                    TareasActivosPreventivosDanios.objects.filter(mantenimiento=mantenimientosactivos).delete()
                    c = 0
                    while c < len(danio):
                        detalle = TareasActivosPreventivosDanios(mantenimiento=mantenimientosactivos,
                                                                 grupos_id=int(MantenimientoGruDanios.objects.get(
                                                                     descripcion=danio[c], grupocategoria_id=id_tipact,
                                                                     status=True).id),
                                                                 estadodanio=daniodes[c])
                        detalle.save(request)
                        c += 1
                    PiezaParteActivosPreventivos.objects.filter(mantenimiento=mantenimientosactivos).delete()
                    counter = 0
                    while counter < len(piezaparte):
                        detalle = PiezaParteActivosPreventivos(mantenimiento=mantenimientosactivos,
                                                               piezaparte_id=int(HdPiezaPartes.objects.get(
                                                                   descripcion=piezaparte[counter], status=True).id),
                                                               descripcion=caracteristica[counter])
                        detalle.save(request)
                        counter += 1
                    log(u'Editó mantenimiento preventivo Activo: %s' % mantenimientosactivos, request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'addtarjeta':
            try:
                f = DetalleTarjetaControlForm(request.POST)
                if f.is_valid():
                    tarjeta = TarjetaControl.objects.get(pk=int(request.POST['id']))
                    detalle = DetalleMantenimiento(tarjeta=tarjeta,
                                                   mantenimientorealizar=f.cleaned_data['mantenimientorealizar'],
                                                   aplicagarantia=f.cleaned_data['aplicagarantia'],
                                                   observacion=f.cleaned_data['observacion'],
                                                   fechaentrega=f.cleaned_data['fechaentrega'],
                                                   estado=1,
                                                   taller=f.cleaned_data['taller'])
                    detalle.save(request)
                    log(u'Adicionó tarjeta: %s' % detalle, request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'ingresodemantenimiento':
            try:
                f = DetalleTarjetaControlForm(request.POST)
                if f.is_valid():
                    detalle = DetalleMantenimiento.objects.get(pk=int(request.POST['id']))
                    detalle.costomanodeobra = f.cleaned_data['costomanodeobra']
                    detalle.mantenimientorealizado = f.cleaned_data['mantenimientorealizado']
                    detalle.aplicagarantia = f.cleaned_data['aplicagarantia']
                    detalle.manodeobra = f.cleaned_data['manodeobra']
                    detalle.repuestos = f.cleaned_data['repuestos']
                    detalle.costomanodereparacion = f.cleaned_data['costomanodereparacion']
                    detalle.facturamanodeobra = f.cleaned_data['facturamanodeobra']
                    detalle.facturareparacion = f.cleaned_data['facturareparacion']
                    detalle.fecharecepcion = f.cleaned_data['fecharecepcion']
                    detalle.observacion = f.cleaned_data['observacion']
                    detalle.estado = 2
                    detalle.save(request)
                    log(u'ingreso de mantenimiento : %s' % detalle, request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'edittarjeta':
            try:
                f = DetalleTarjetaControlForm(request.POST)
                if f.is_valid():
                    detalle = DetalleMantenimiento.objects.get(pk=int(request.POST['id']))
                    detalle.fechaentrega = f.cleaned_data['fechaentrega']
                    detalle.mantenimientorealizar = f.cleaned_data['mantenimientorealizar']
                    detalle.taller = f.cleaned_data['taller']
                    detalle.observacion = f.cleaned_data['observacion']
                    detalle.estado = 2
                    detalle.save(request)
                    log(u'Editó tarjeta: %s' % detalle, request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'addtraslado':
            try:
                f = TrasladoMantenimientoForm(request.POST)
                if f.is_valid():
                    traslado = TrasladoMantenimiento(departamentosolicita=f.cleaned_data['departamentosolicita'],
                                                     fecha=datetime.now(),
                                                     asistentelogistica=f.cleaned_data['asistentelogistica'],
                                                     usuariobienes=f.cleaned_data['usuariobienes'],
                                                     taller=f.cleaned_data['taller'],
                                                     administradorcontrato=f.cleaned_data['administradorcontrato'],
                                                     observacion=f.cleaned_data['observacion'])
                    traslado.save(request)
                    log(u'Adicionó traslado: %s' % traslado, request, "add")
                    cursor = connection.cursor()
                    cursor.execute(
                        "INSERT INTO sagest_detalletrasladomantenimiento (usuario_creacion_id, fecha_creacion, codigotraslado_id, activo_id, observacion, status, seleccionado) "
                        "SELECT %s, now(), %s, id, '', true, false from sagest_activofijo WHERE statusactivo=1 AND responsable_id=%s;",
                        [request.user.id, traslado.id, traslado.usuariobienes.id])
                    return JsonResponse({"result": "ok", "id": traslado.id})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'addbaja':
            try:
                f = BajaActivoForm(request.POST)
                if f.is_valid():
                    if f.cleaned_data['fechaoficio'] > f.cleaned_data['fecha']:
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"La fecha de oficio no debe ser mayor a la fecha de baja"})
                    baja = BajaActivo(fecha=f.cleaned_data['fecha'],
                                      tiposolicitud=f.cleaned_data['tiposolicitud'],
                                      solicitante=f.cleaned_data['solicitante'],
                                      tipobaja=f.cleaned_data['tipobaja'],
                                      ubicacionbienentrega=f.cleaned_data['ubicacionbienentrega'],
                                      usuariobienentrega=f.cleaned_data['usuariobienentrega'],
                                      custodioentrega=f.cleaned_data['custodioentrega'],
                                      # usuariorecibe=f.cleaned_data['usuariorecibe'],
                                      oficio=f.cleaned_data['oficio'],
                                      # cargorecibe=f.cleaned_data['cargorecibe'],
                                      fechaoficio=f.cleaned_data['fechaoficio'],
                                      experto=f.cleaned_data['experto'],
                                      contadorper=f.cleaned_data['contadorper'],
                                      usuarioejecuta=f.cleaned_data['usuarioejecuta'],
                                      observacion=f.cleaned_data['observacion'])
                    baja.save(request)
                    log(u'Adicionó baja: %s' % baja, request, "add")
                    cursor = connection.cursor()
                    cursor.execute(
                        "INSERT INTO sagest_detallebajaactivo (usuario_creacion_id, fecha_creacion, codigobaja_id, activo_id, status, seleccionado) "
                        "SELECT %s, now(), %s, id, True, False from sagest_activofijo WHERE statusactivo=1 AND responsable_id=%s AND ubicacion_id=%s AND custodio_id=%s;",
                        [request.user.id, baja.id, baja.usuariobienentrega.id, baja.ubicacionbienentrega.id,
                         baja.custodioentrega.id])
                    return JsonResponse({"result": "ok", "id": baja.id})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'addtipobaja':
            try:
                f = TipoBajaForm(request.POST)
                if f.is_valid():
                    if TipoBaja.objects.filter(nombre=f.cleaned_data['nombre']).exists():
                        return JsonResponse({"result": "bad", "mensaje": u"Registro ya existe"})
                    tipobaja = TipoBaja(nombre=f.cleaned_data['nombre'])
                    tipobaja.save(request)
                    log(u'Adicionó Tipo baja: %s' % tipobaja, request, "add")
                    return JsonResponse({"result": "ok", "id": tipobaja.id})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'edittraslado':
            try:
                f = TrasladoMantenimientoForm(request.POST)
                if f.is_valid():
                    traslado = TrasladoMantenimiento.objects.get(pk=int(request.POST['id']))
                    traslado.observacion = f.cleaned_data['observacion']
                    traslado.save(request)
                    return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'finalizarcons':
            try:
                constatacion = ConstatacionFisica.objects.get(pk=int(request.POST['id']))
                detalle = DetalleConstatacionFisica.objects.filter(status=True, codigoconstatacion=constatacion)
                constatacion.estado = 2
                constatacion.numero = constatacion.secuencial_acta()
                constatacion.fechafin = datetime.now()
                constatacion.usuariofinaliza = persona
                constatacion.save(request)
                for det in detalle:
                    if det.estadoactual and not det.activo.estado == det.estadoactual:
                        det.activo.estado = det.estadoactual
                        det.activo.save(request)
                        log(u'Cambió estado de activo por constatación: %s' % det.activo, request, "edit")
                log(u'Finalizar constatacion: %s' % constatacion, request, "edit")
                url_archivo = generar_acta_constatacion(request, constatacion)
                acta = ActaConstatacion(constatacion=constatacion,
                                        persona=persona,
                                        archivo=url_archivo)
                acta.save(request)
                log(u'Creo historial de acta de constatación: %s' % acta, request, "add")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": f"Error: {ex}({sys.exc_info()[-1].tb_lineno})"})

        elif action == 'finalizarconstatacion':
            try:
                from django.db.models import F
                constatacion = ConstatacionFisica.objects.get(pk=encrypt_id(request.POST['id']))
                fecha = request.POST['fecha']
                hora = hoy.time()
                # Parsea la fecha en formato de cadena a un objeto de fecha
                fecha_obj = datetime.strptime(fecha, '%Y-%m-%d').date()
                # Combina la fecha y la hora para crear un objeto de fecha y hora
                fecha_hora = datetime.combine(fecha_obj, hora)
                # detalle = DetalleConstatacionFisica.objects.filter(status=True, codigoconstatacion=constatacion)
                constatacion.estado = 2
                constatacion.numero = constatacion.secuencial_acta()
                constatacion.fechafin = fecha_hora
                constatacion.usuariofinaliza = persona
                constatacion.save(request)

                detalle = constatacion.detalle_constatacion().filter(estadoactual__isnull=False).exclude(activo__estado=F('estadoactual'))
                # detalle.update(activo__estado=F('estadoactual'))
                for det in detalle:
                    activo = det.activo
                    activo.estado = det.estadoactual
                    activo.save(request)
                    log(u'Cambió estado de activo por constatación: %s' % det.activo, request, "edit")
                log(u'Finalizar constatacion: %s' % constatacion, request, "edit")
                if len(constatacion.detalle_constatacion()) > 1000:
                    data['eConstatacion']=constatacion
                    data['persona']=persona
                    titulo = f'Generando acta de constatación de activos de {constatacion.usuariobienes}'
                    noti = Notificacion(cuerpo='Se inicializo la creación d acta de activos de constatacion de bienes en segundo plano por exceso de registros',
                                        titulo=titulo, destinatario=persona,
                                        url='',
                                        prioridad=1, app_label='SGA-SAGEST',
                                        fecha_hora_visible=datetime.now() + timedelta(days=1), tipo=2,
                                        en_proceso=True)
                    noti.save(request)
                    generar_acta_constatacion_background(request=request, data=data, notif=noti.pk).start()
                else:
                    pdf = generar_acta_constatacion_reportlab(request, constatacion)
                    acta = ActaConstatacion(constatacion=constatacion,
                                            persona=persona,
                                            archivo=pdf)
                    acta.save(request)
                    log(u'Creo historial de acta de constatación: %s' % acta, request, "add")
                return JsonResponse({"result": False})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": f"Error: {ex}({sys.exc_info()[-1].tb_lineno})"})

        elif action == 'generaractaconstatacion':
            try:
                constatacion = ConstatacionFisica.objects.get(id=encrypt_id(request.POST['id']))
                constatacion.estadoacta = 1
                constatacion.save(request)
                actas = ActaConstatacion.objects.filter(status=True, constatacion=constatacion)
                actas.update(status=False)
                if len(constatacion.detalle_constatacion()) > 1500:
                    data['eConstatacion'] = constatacion
                    data['persona'] = persona
                    titulo = f'Generando acta de constatación de activos de {constatacion.usuariobienes}'
                    noti = Notificacion(cuerpo='Se inicializo la creación d acta de activos de constatacion de bienes en segundo plano por exceso de registros',
                                        titulo=titulo, destinatario=persona,
                                        url='',
                                        prioridad=1, app_label='SGA-SAGEST',
                                        fecha_hora_visible=datetime.now() + timedelta(days=1), tipo=2,
                                        en_proceso=True)
                    noti.save(request)
                    generar_acta_constatacion_background(request=request, data=data, notif=noti.pk).start()
                    return JsonResponse({'result': 'ok','showSwal':True,'mensaje':'Acta de constatación generandose en segundo plano'})
                else:
                    pdf = generar_acta_constatacion_reportlab(request, constatacion)
                    acta = ActaConstatacion(constatacion=constatacion,
                                            persona=persona,
                                            archivo=pdf)
                    acta.save(request)
                    log(u'Creo historial de acta de constatación: %s' % acta, request, "add")
                    return JsonResponse({'result': 'ok', })
            except Exception as ex:
                return JsonResponse({'result': 'bad', 'mensaje':f'Error: {ex}'})

        elif action == 'revertirestadoacta':
            try:
                constatacion = ConstatacionFisica.objects.get(id=encrypt_id(request.POST['id']))
                constatacion.estadoacta = 1
                constatacion.save(request)
                actas = ActaConstatacion.objects.filter(status=True, constatacion=constatacion)
                actas.update(status=False)
                log(u'Cambio estado de constatación: %s' % constatacion, request, "edit")
                return JsonResponse({'result': 'ok', })
            except Exception as ex:
                return JsonResponse({'result': 'bad', 'mensaje':f'Error: {ex}'})

        elif action == 'confirmarexpo':
            try:
                exportacion = ExportacionesActivos.objects.get(pk=int(request.POST['id']))
                exportacion.estado = 2
                exportacion.save(request)
                lista = exportacion.activos.all().values_list('id', flat=True)
                ActivoFijo.objects.filter(id__in=lista).update(subidogobierno=True)
                log(u'Finalizar exportacion: %s' % exportacion, request, "edit")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'finalizartraslado':
            try:
                traslado = TrasladoMantenimiento.objects.get(pk=int(request.POST['id']))
                traslado.estado = 2
                traslado.save(request)
                secuencia = secuencia_activos(request)
                if not traslado.numero:
                    secuencia.numeromantenimiento += 1
                    secuencia.save(request)
                    traslado.numero = secuencia.numeromantenimiento
                    traslado.save(request)
                traslado.detalletrasladomantenimiento_set.filter(seleccionado=False).update(status=False)
                for detalle in traslado.detalletrasladomantenimiento_set.filter(seleccionado=True):
                    detallematenimiento = DetalleMantenimiento(tarjeta=detalle.activo.mi_tarjeta_control(),
                                                               detalletrasladomantenimiento=detalle,
                                                               taller=traslado.taller,
                                                               fechaentrega=traslado.fecha,
                                                               mantenimientorealizar=detalle.observacion)
                    detallematenimiento.save(request)
                log(u'Finalizar traslado: %s' % traslado, request, "edit")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'finalizartraspaso':
            try:
                traspaso = TraspasoActivo.objects.get(pk=int(request.POST['id']))
                if not traspaso.detalletraspasoactivo_set.filter(seleccionado=True).exists():
                    return JsonResponse({"result": "bad", "mensaje": u"Debe seleccionar un activo."})
                traspaso.estado = 2
                traspaso.save(request)
                secuencia = secuencia_activos(request)
                if not traspaso.numero:
                    secuencia.numerotraspaso += 1
                    secuencia.save(request)
                    traspaso.numero = secuencia.numerotraspaso
                traspaso.normativatraspaso = secuencia.normativatraspaso
                traspaso.save(request)
                traspaso.detalletraspasoactivo_set.filter(seleccionado=False).update(status=False)
                traspaso.totalbienes = traspaso.cantidad_seleccionados()
                traspaso.save(request)
                solicitudtraspasoactivo = SolicitudTraspasoActivos.objects.filter(status=True,
                                                                                  traspasoactivofijo=traspaso, estado=8)
                if solicitudtraspasoactivo:
                    solicitudtraspasoactivo[0].estado = 11
                    solicitudtraspasoactivo[0].puedefirmar = True
                    solicitudtraspasoactivo[0].save(request)
                    log(u'Solicitud traspaso realizado: %s' % solicitudtraspasoactivo[0], request, "act")
                    seguimientosolicitudtraspaso = SeguimientoSolicitudTraspaso(
                        solicitudtraspaso=solicitudtraspasoactivo[0], estado=9)
                    seguimientosolicitudtraspaso.save(request)
                    seguimientosolicitudtraspaso = SeguimientoSolicitudTraspaso(
                        solicitudtraspaso=solicitudtraspasoactivo[0], estado=11)
                    seguimientosolicitudtraspaso.save(request)
                    log(u'Traspaso pendiente firmar: %s' % seguimientosolicitudtraspaso, request, "add")
                    reporte = Reporte.objects.get(id=211)
                    cambiaruta = 0
                    isQR = False
                    codigo = None
                    certificado = None
                    base_url = request.META['HTTP_HOST']
                    if reporte.archivo:
                        base_url = request.META['HTTP_HOST']
                        d = datetime.now()
                        pdfname = reporte.nombre + d.strftime('%Y%m%d_%H%M%S')
                        tipo = 'pdf'
                        paRequest = {
                            'id': traspaso.id,
                            'imp_logo': True,
                            'imp_encabezado': True,
                            'imp_fecha': True,
                            'imp_membretada': False,
                            'url_qr': unicode(base_url + "/".join(
                                [MEDIA_URL, 'documentos', 'userreports', elimina_tildes(request.user.username),
                                 pdfname + "." + tipo]))
                        }
                        d = run_report_v1(reporte=reporte, tipo=tipo, paRequest=paRequest, request=request)
                        if not d['isSuccess']:
                            raise NameError(d['mensaje'])
                        else:
                            archivoobtenido = None
                            archivogeneradotraspaso = d['data']['reportfile']
                            data['url_archivo'] = rutaarchivoparafirmar = '{}{}'.format(dominio_sistema,
                                                                                        archivogeneradotraspaso)
                            url_archivo = (SITE_STORAGE + archivogeneradotraspaso).replace('\\', '/')
                            url_archivo = (url_archivo).replace('//', '/')
                            _name = generar_nombre(f'archivotraspasoactivo{request.user.username}_{traspaso.id}_',
                                                   'generado')
                            folder = os.path.join(SITE_STORAGE, 'media', 'traspasoactivofirma', '')
                            if not os.path.exists(folder):
                                os.makedirs(folder)
                            folder_save = os.path.join('traspasoactivofirma', '').replace('\\', '/')
                            url_file_generado = f'{folder_save}{_name}.pdf'
                            ruta_creacion = SITE_STORAGE
                            ruta_creacion = ruta_creacion.replace('\\', '/')
                            shutil.copy(url_archivo, ruta_creacion + '/media/' + url_file_generado)
                            traspaso.traspasoactivofirma = url_file_generado
                            traspaso.save(request)
                for detalle in traspaso.detalletraspasoactivo_set.filter(status=True):
                    activo = detalle.activo
                    responsable = traspaso.usuariobienrecibe if traspaso.usuariobienrecibe else activo.responsable
                    responsableanterior = activo.responsable
                    activo.actualiza_responsable_directo(traspaso.custodiobienrecibe, traspaso.ubicacionbienrecibe,
                                                         traspaso.usuariobienrecibe if traspaso.usuariobienrecibe else activo.responsable)
                    solicitud = SolicitudActivos.objects.filter(status=True, activo_id=activo.pk,
                                                                responsableasignacion=responsable,
                                                                estado=2)
                    if solicitud:
                        historial = HistorialTraspaso(activofijo_id=activo.id, responsableanterior=responsableanterior,
                                                      responsableasignado=responsable,
                                                      fechasolicitud=solicitud[0].fechasolicitud,
                                                      fechaasignacion=datetime.now().date())
                        solicitud[0].estado = 7
                        solicitud[0].save(request)
                        historial.save(request)
                        log(u'Solicitud traspaso finalizado: %s' % solicitud[0], request, "add")
                    activo.save(request)
                log(u'Finalizar traspaso: %s' % traspaso, request, "edit")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'abrirtraspaso':
            try:
                traspaso = TraspasoActivo.objects.get(pk=int(request.POST['id']))
                traspaso.estado = 1
                traspaso.save(request)
                traspaso.detalletraspasoactivo_set.all().update(status=True)
                for detalle in traspaso.detalletraspasoactivo_set.all():
                    activo = detalle.activo
                    activo.actualiza_responsable()
                    activo.save(request)
                log(u'Abrió traspaso: %s' % traspaso, request, "edit")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'actualizamasivo':
            try:
                activos = ActivoFijo.objects.filter(responsable__isnull=True, status=True)
                for act in activos:
                    traspaso = act.detalletraspasoactivo_set.filter(status=True).order_by(
                        '-codigotraspaso__fecha').last()
                    act.custodio = traspaso.codigotraspaso.custodiobienrecibe
                    act.ubicacion = traspaso.codigotraspaso.ubicacionbienrecibe
                    act.responsable = traspaso.codigotraspaso.usuariobienrecibe
                    act.save(request)
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'eliminartraspaso':
            try:
                traspaso = TraspasoActivo.objects.get(pk=int(request.POST['id']))
                traspaso.status = False
                traspaso.save(request)
                traspaso.detalletraspasoactivo_set.all().update(status=False)
                log(u'Eliminar traspaso: %s' % traspaso, request, "delete")
                traspasonorealizado = SolicitudTraspasoActivos.objects.filter(traspasoactivofijo=traspaso, estado=8,
                                                                              status=True)
                if traspasonorealizado:
                    traspasonorealizado[0].estado = 10
                    traspasonorealizado[0].save()
                    log(u'Traspaso no realizado: %s' % traspaso, request, "act")
                    seguimiento = SeguimientoSolicitudTraspaso(solicitudtraspaso=traspasonorealizado[0], estado=10)
                    seguimiento.save(request)
                    log(u'Seguimiento traspaso no realizado: %s' % seguimiento, request, "add")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'deleteexpo':
            try:
                exportacion = ExportacionesActivos.objects.get(pk=int(request.POST['id']))
                exportacion.delete()
                log(u'Eliminar exportacion: %s' % exportacion, request, "delete")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'deleteimpo':
            try:
                exportacion = ArchivoActivoFijo.objects.get(pk=int(request.POST['id']))
                exportacion.delete()
                log(u'Eliminar importacion: %s' % exportacion, request, "delete")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})
        elif action == 'delete':
            try:
                with transaction.atomic():
                    instancia = CatalogoBien.objects.get(pk=int(request.POST['id']))
                    instancia.clasificado = False
                    instancia.save(request)
                    log(u'Elimino Catalogo: %s' % instancia, request, "delete")
                    res_json = {"error": False}
            except Exception as ex:
                res_json = {'error': True, "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)
        elif action == 'eliminartraslado':
            try:
                traslado = TrasladoMantenimiento.objects.get(pk=int(request.POST['id']))
                traslado.status = False
                traslado.save(request)
                traslado.detalletrasladomantenimiento_set.all().update(status=False)
                log(u'Eliminar traspaso: %s' % traslado, request, "delete")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'eliminarconstatacion':
            try:
                constatacion = ConstatacionFisica.objects.get(pk=int(request.POST['id']))
                constatacion.status = False
                constatacion.save(request)
                constatacion.detalleconstatacionfisica_set.all().update(status=False)
                constatacion.detallenoidentificado_set.all().update(status=False)
                log(u'Eliminar traspaso: %s' % constatacion, request, "delete")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'eliminarbaja':
            try:
                baja = BajaActivo.objects.get(pk=int(request.POST['id']))
                baja.status = False
                baja.save(request)
                baja.detallebajaactivo_set.all().update(status=False)
                log(u'Eliminar baja: %s' % baja, request, "delete")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'eliminartipobaja':
            try:
                tipobaja = TipoBaja.objects.get(pk=int(request.POST['id']))
                log(u'Eliminar Tipo baja: %s' % tipobaja, request, "delete")
                tipobaja.delete()
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'finalizarbaja':
            try:
                baja = BajaActivo.objects.get(pk=int(request.POST['id']))
                baja.estado = 2
                baja.save(request)
                secuencia = secuencia_activos(request)
                if not baja.numero:
                    secuencia.numerobajaactivo += 1
                    secuencia.save(request)
                    baja.numero = secuencia.numerobajaactivo
                    baja.save(request)
                baja.detallebajaactivo_set.filter(seleccionado=False).update(status=False)
                for detalle in baja.detallebajaactivo_set.filter(seleccionado=True):
                    activo = detalle.activo
                    activo.depreciar_activo(baja.fecha, request)
                    activo.statusactivo = 2
                    activo.save(request)
                log(u'Finalizar traslado: %s' % baja, request, "edit")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'finalizarmantenimiento':
            try:
                mantenimiento = DetalleTrasladoMantenimiento.objects.get(pk=int(request.POST['id']))
                detalle = mantenimiento.detallemantenimiento_set.all()[0]
                detalle.estado = 2
                detalle.save(request)
                log(u'Finalizar tarjeta de control: %s' % detalle, request, "edit")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'finalizaracta':
            try:
                acta = TraspasoActivo.objects.get(pk=int(request.POST['id']))
                acta.estado = 2
                acta.save(request)
                log(u'Finalizar acta de entrega: %s' % acta, request, "edit")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'cambiousuario':
            try:
                f = ConstatacionForm(request.POST)
                if f.is_valid():
                    usuario = Persona.objects.get(id=int(f.cleaned_data['usuariobienes']))
                    constatacion = ConstatacionFisica.objects.get(pk=int(request.POST['id']))
                    constatacion.detalleconstatacionfisica_set.all().delete()
                    constatacion.usuariobienes = usuario
                    constatacion.ubicacionbienes = f.cleaned_data['ubicacionbienes']
                    constatacion.save(request)
                    cursor = connection.cursor()
                    cursor.execute(
                        "INSERT INTO sagest_detalleconstatacionfisica (status, usuario_creacion_id, fecha_creacion, codigoconstatacion_id, activo_id, perteneceusuario, requieretraspaso, estadooriginal_id, estadoactual_id, encontrado, enuso, observacion) "
                        "SELECT true, %s, now(), %s, id, true, false, estado_id, estado_id, false, false, ''  from sagest_activofijo WHERE statusactivo=1 AND responsable_id=%s AND ubicacion_id=%s;",
                        [request.user.id, constatacion.id, constatacion.usuariobienes.id,
                         constatacion.ubicacionbienes.id])
                    return JsonResponse({"result": "ok", "id": constatacion.id})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'addconstatacion':
            try:
                f = ConstatacionForm(request.POST)
                if f.is_valid():
                    secuencia = secuencia_activos(request)
                    secuencia.numeroconstatacion += 1
                    secuencia.save(request)
                    constatacion = ConstatacionFisica(fechainicio=datetime.now(),
                                                      estado=1,
                                                      numero=secuencia.numeroconstatacion,
                                                      normativaconstatacion=secuencia.normativaconstatacion,
                                                      usuariobienes_id=f.cleaned_data['usuariobienes'],
                                                      ubicacionbienes=f.cleaned_data['ubicacionbienes'],
                                                      observacion=f.cleaned_data['observacion'])
                    constatacion.save(request)
                    log(u'Adicionó contratación Fisica: %s' % constatacion, request, "add")
                    cursor = connection.cursor()
                    cursor.execute(
                        "INSERT INTO sagest_detalleconstatacionfisica (usuario_creacion_id, fecha_creacion, codigoconstatacion_id, activo_id, perteneceusuario, requieretraspaso, estadooriginal_id, estadoactual_id, encontrado, enuso, status, observacion) "
                        "SELECT %s, now(), %s, id, true, false, estado_id, estado_id, false, false, true, '' from sagest_activofijo WHERE statusactivo=1 AND responsable_id=%s AND ubicacion_id=%s;",
                        [request.user.id, constatacion.id, constatacion.usuariobienes.id,
                         constatacion.ubicacionbienes.id])
                    return JsonResponse({"result": "ok", "id": constatacion.id})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'addtraspaso':
            try:
                f = TraspasoActivoForm(request.POST)
                if f.is_valid():
                    entrega = f.cleaned_data['usuariobienentrega']
                    if f.cleaned_data['fechaoficio'] > f.cleaned_data['fecha']:
                        return JsonResponse({"result": "bad",
                                             "mensaje": u"La fecha de oficio no debe ser mayor a la fecha de traspaso"})
                    if f.cleaned_data['custodiobienrecibe'] == f.cleaned_data['custodiobienentrega'] and f.cleaned_data[
                        'usuariobienentrega'] == f.cleaned_data['usuariobienrecibe'] and f.cleaned_data[
                        'ubicacionbienentrega'] == f.cleaned_data['ubicacionbienrecibe']:
                        return JsonResponse({"result": "bad",
                                             "mensaje": u"Está tratando de realizar un traspasos. con datos de entrega y recepción similares"})
                    traspaso = TraspasoActivo(custodiobienrecibe=f.cleaned_data['custodiobienrecibe'],
                                              usuariobienentrega=f.cleaned_data['usuariobienentrega'],
                                              custodiobienentrega=f.cleaned_data['custodiobienentrega'],
                                              usuariobienrecibe=f.cleaned_data['usuariobienrecibe'],
                                              ubicacionbienentrega=f.cleaned_data['ubicacionbienentrega'],
                                              responsablebienes_id=1204,
                                              ubicacionbienrecibe=f.cleaned_data['ubicacionbienrecibe'],
                                              fecha=f.cleaned_data['fecha'],
                                              solicitante=f.cleaned_data['solicitante'],
                                              fechaoficio=f.cleaned_data['fechaoficio'],
                                              observacion=f.cleaned_data['observacion'],
                                              oficio=f.cleaned_data['oficio'],
                                              tiposolicitud=f.cleaned_data['tiposolicitud'],
                                              tipotraspaso=1,
                                              tipo=2)
                    traspaso.save(request)
                    log(u'Adicionó traspaso activo: %s' % traspaso, request, "add")
                    cursor = connection.cursor()
                    cursor.execute(
                        "INSERT INTO sagest_detalletraspasoactivo (usuario_creacion_id, fecha_creacion, codigotraspaso_id, activo_id, status, historico, seleccionado) "
                        "SELECT %s, now(), %s, id, true, false, false from sagest_activofijo WHERE procesobaja=false AND statusactivo=1 AND responsable_id=%s AND ubicacion_id=%s AND custodio_id=%s;",
                        [request.user.id, traspaso.id, traspaso.usuariobienentrega.id, traspaso.ubicacionbienentrega.id,
                         traspaso.custodiobienentrega.id])
                    return JsonResponse({"result": "ok", "id": traspaso.id})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'addtraspasocustodio':
            try:
                f = TraspasoActivoCustodioForm(request.POST)
                if f.is_valid():
                    if f.cleaned_data['fechaoficio'] > f.cleaned_data['fecha']:
                        return JsonResponse({"result": "bad",
                                             "mensaje": u"La fecha de oficio no debe ser mayor a la fecha de traspaso"})
                    if f.cleaned_data['custodiobienrecibe'] == f.cleaned_data['custodiobienentrega']:
                        return JsonResponse({"result": "bad",
                                             "mensaje": u"Está tratando de realizar un traspasos. con datos de entrega y recepción similares"})
                    traspaso = TraspasoActivo(custodiobienrecibe=f.cleaned_data['custodiobienrecibe'],
                                              custodiobienentrega=f.cleaned_data['custodiobienentrega'],
                                              ubicacionbienentrega=f.cleaned_data['ubicacionbienentrega'],
                                              ubicacionbienrecibe=f.cleaned_data['ubicacionbienentrega'],
                                              responsablebienes_id=1204,
                                              fecha=f.cleaned_data['fecha'],
                                              solicitante=f.cleaned_data['solicitante'],
                                              fechaoficio=f.cleaned_data['fechaoficio'],
                                              observacion=f.cleaned_data['observacion'],
                                              oficio=f.cleaned_data['oficio'],
                                              tiposolicitud=f.cleaned_data['tiposolicitud'],
                                              tipotraspaso=2,
                                              tipo=2)
                    traspaso.save(request)
                    log(u'Adicionó traspaso custio de  activo: %s' % traspaso, request, "add")
                    cursor = connection.cursor()
                    cursor.execute(
                        "INSERT INTO sagest_detalletraspasoactivo (usuario_creacion_id, fecha_creacion, codigotraspaso_id, activo_id, historico, seleccionado, status) "
                        "SELECT %s, now(), %s, id, false, false, true from sagest_activofijo WHERE statusactivo=1 AND custodio_id=%s AND ubicacion_id=%s;",
                        [request.user.id, traspaso.id, traspaso.custodiobienentrega.id,
                         traspaso.ubicacionbienentrega.id])
                    return JsonResponse({"result": "ok", "id": traspaso.id})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'adddirector':
            try:
                f = DirectorResponsableBajaForm(request.POST)
                if f.is_valid():
                    if f.cleaned_data['fechafin']:
                        if f.cleaned_data['fechainicio'] > f.cleaned_data['fechafin']:
                            return JsonResponse(
                                {"result": "bad", "mensaje": u"La fecha inicio no puede ser mayor que la fecha fin "})

                    director = DirectorResponsableBaja(
                        responsable_id=f.cleaned_data['responsable'],
                        cargo=f.cleaned_data['cargo'],
                        fechainicio=f.cleaned_data['fechainicio'],
                    )
                    if f.cleaned_data['actual']:
                        director.actual = True
                        if DirectorResponsableBaja.objects.filter(status=True, actual=True).exists():
                            cerrar = DirectorResponsableBaja.objects.get(status=True, actual=True)
                            cerrar.fechafin = datetime.now().date()
                            cerrar.actual = False
                            cerrar.save()
                    else:
                        director.fechafin = f.cleaned_data['fechafin']
                    director.save(request)
                    log(u'Adicionó director: %s' % director, request, "add")
                    return JsonResponse({"result": False}, safe=False)
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        elif action == 'editdirector':
            try:
                f = DirectorResponsableBajaForm(request.POST)
                director = DirectorResponsableBaja.objects.get(pk=int(request.POST['id']))
                if f.is_valid():
                    if f.cleaned_data['fechafin']:
                        if f.cleaned_data['fechainicio'] > f.cleaned_data['fechafin']:
                            return JsonResponse(
                                {"result": "bad", "mensaje": u"La fecha inicio no puede ser mayor que la fecha fin "})

                    director.fechainicio = f.cleaned_data['fechainicio']
                    director.cargo = f.cleaned_data['cargo']
                    if f.cleaned_data['actual']:
                        director.actual = True
                        if DirectorResponsableBaja.objects.filter(status=True, actual=True).exists():
                            cerrar = DirectorResponsableBaja.objects.get(status=True, actual=True)
                            cerrar.fechafin = datetime.now().date()
                            cerrar.actual = False
                            cerrar.save()
                    else:
                        director.fechafin = f.cleaned_data['fechafin']
                    director.save(request)
                    log(u'Adicionó director: %s' % director, request, "add")
                    return JsonResponse({"result": False}, safe=False)
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        elif action == 'activosusuarioconstatacion':
            try:
                data = {}
                data['constatacion'] = constatacion = ConstatacionFisica.objects.get(pk=int(request.POST['id']))
                detalle = constatacion.detalleconstatacionfisica_set.filter(perteneceusuario=True)
                detallebaja = detalle.filter(activo__procesobaja=True)
                detallebaja.update(encontrado=True)
                for det in detallebaja:
                    det.estadoactual = det.activo.estado
                    det.save()

                pagina = 1
                paging = MiPaginador(detalle, 10)
                p = 1
                try:
                    if 'page' in request.POST:
                        p = int(request.POST['page'])
                    page = paging.page(p)
                except:
                    page = paging.page(1)
                data['paging'] = paging
                data['rangospaging'] = paging.rangos_paginado(p)
                data['page'] = page
                data['detalles'] = page.object_list
                data['usuario'] = request.user
                template = get_template("af_activofijo/detalleconstatacion.html")
                json_content = template.render(data)
                return JsonResponse(
                    {"result": "ok", 'data': json_content, 'contador': constatacion.contadores_pertenece()})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'activostraspaso':
            try:
                data = {}
                data['traspaso'] = traspaso = TraspasoActivo.objects.get(pk=int(request.POST['id']))
                detalle = traspaso.detalletraspasoactivo_set.all()
                pagina = 1
                paging = MiPaginador(detalle, 10)
                p = 1
                try:
                    if 'page' in request.POST:
                        p = int(request.POST['page'])
                    page = paging.page(p)
                except:
                    page = paging.page(1)
                data['paging'] = paging
                data['rangospaging'] = paging.rangos_paginado(p)
                data['page'] = page
                data['detalles'] = page.object_list
                data['usuario'] = request.user
                template = get_template("af_activofijo/detalletraspaso.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'data': json_content, 'contador': traspaso.contador_traspaso()})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'activosacta':
            try:
                data = {}
                data['acta'] = acta = TraspasoActivo.objects.get(pk=int(request.POST['id']))
                detalle = acta.detalletraspasoactivo_set.all()
                pagina = 1
                paging = MiPaginador(detalle, 10)
                p = 1
                try:
                    if 'page' in request.POST:
                        p = int(request.POST['page'])
                    page = paging.page(p)
                except:
                    page = paging.page(1)
                data['paging'] = paging
                data['rangospaging'] = paging.rangos_paginado(p)
                data['page'] = page
                data['detalles'] = page.object_list
                data['usuario'] = request.user
                template = get_template("af_activofijo/detalleacta.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'data': json_content})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'activosbaja':
            try:
                data = {}
                data['baja'] = baja = BajaActivo.objects.get(pk=int(request.POST['id']))
                detalle = baja.detallebajaactivo_set.all()
                pagina = 1
                paging = MiPaginador(detalle, 10)
                p = 1
                try:
                    if 'page' in request.POST:
                        p = int(request.POST['page'])
                    page = paging.page(p)
                except:
                    page = paging.page(1)
                data['paging'] = paging
                data['rangospaging'] = paging.rangos_paginado(p)
                data['page'] = page
                data['detalles'] = page.object_list
                data['usuario'] = request.user
                template = get_template("af_activofijo/detallebajaactivos.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'data': json_content, 'contador': baja.contador()})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'activostraslado':
            try:
                data = {}
                data['traslado'] = traslado = TrasladoMantenimiento.objects.get(pk=int(request.POST['id']))
                detalle = traslado.detalletrasladomantenimiento_set.all()
                pagina = 1
                paging = MiPaginador(detalle, 10)
                p = 1
                try:
                    if 'page' in request.POST:
                        p = int(request.POST['page'])
                    page = paging.page(p)
                except:
                    page = paging.page(1)
                data['paging'] = paging
                data['rangospaging'] = paging.rangos_paginado(p)
                data['page'] = page
                data['detalles'] = page.object_list
                data['usuario'] = request.user
                template = get_template("af_activofijo/activostraslado.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'data': json_content, 'contador': traslado.contador()})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'activosotrousuarioconstatacion':
            try:
                data = {}
                constatacion = ConstatacionFisica.objects.get(pk=int(request.POST['id']))
                detalle = constatacion.detalleconstatacionfisica_set.filter(perteneceusuario=False).order_by(
                    'activo__codigointerno', 'activo__codigogobierno')
                pagina = 1
                paging = MiPaginador(detalle, 10)
                p = 1
                try:
                    if 'page' in request.POST:
                        p = int(request.POST['page'])
                    page = paging.page(p)
                except:
                    page = paging.page(1)
                data['paging'] = paging
                data['rangospaging'] = paging.rangos_paginado(p)
                data['page'] = page
                data['detalles'] = page.object_list
                data['usuario'] = request.user
                template = get_template("af_activofijo/detalleotrosconstatacion.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'data': json_content, 'contador': constatacion.contadores_otros()})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'activosnoidentificadosconstatacion':
            try:
                data = {}
                constatacion = ConstatacionFisica.objects.get(pk=int(request.POST['id']))
                detalle = constatacion.detallenoidentificado_set.all().order_by('codigobarra')
                pagina = 1
                paging = MiPaginador(detalle, 10)
                p = 1
                try:
                    if 'page' in request.POST:
                        p = int(request.POST['page'])
                    page = paging.page(p)
                except:
                    page = paging.page(1)
                data['paging'] = paging
                data['rangospaging'] = paging.rangos_paginado(p)
                data['page'] = page
                data['detalles'] = page.object_list
                data['usuario'] = request.user
                template = get_template("af_activofijo/detalleniconstatacion.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'data': json_content, 'contador': constatacion.contadores_ni()})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'marcarencontrado':
            try:
                detalle = DetalleConstatacionFisica.objects.get(pk=int(request.POST['id']))
                encontrado = request.POST['valencontrado'] == u'true'
                if encontrado:
                    detalle.encontrado = True
                    detalle.enuso = True
                else:
                    detalle.encontrado = False
                    detalle.enuso = False
                    detalle.requieretraspaso = False
                detalle.save(request)
                return JsonResponse(
                    {"result": "ok", "reload": 'False', 'contador': detalle.codigoconstatacion.contadores_pertenece()})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos"})

        elif action == 'marcartodosencontrado':
            try:
                constatacion = ConstatacionFisica.objects.get(pk=int(request.POST['id']))
                detalle = constatacion.detalleconstatacionfisica_set.all()

                if(detalle.filter(encontrado=True).count() < detalle.count()):
                    detalle.filter(encontrado=False).update(encontrado=True, enuso = True)
                # if constatacion.detalleconstatacionfisica_set.all()[0].encontrado == True:
                else:
                    detalle.filter(encontrado=True).update(encontrado=False, enuso=False, requieretraspaso=False)
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos"})

        elif action == 'marcarenuso':
            try:
                detalle = DetalleConstatacionFisica.objects.get(pk=int(request.POST['id']))
                detalle.enuso = request.POST['valenuso'] == u'true'
                detalle.save(request)
                return JsonResponse(
                    {"result": "ok", "reload": 'False', 'contador': detalle.codigoconstatacion.contadores_pertenece()})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos"})

        elif action == 'marcarenusootro':
            try:
                detalle = DetalleConstatacionFisica.objects.get(pk=int(request.POST['id']))
                detalle.enuso = request.POST['valenuso'] == u'true'
                detalle.save(request)
                return JsonResponse(
                    {"result": "ok", "reload": 'False', 'contador': detalle.codigoconstatacion.contadores_otros()})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos"})

        elif action == 'marcartodostraslado':
            try:
                traslado = TrasladoMantenimiento.objects.get(pk=int(request.POST['id']))
                estado = request.POST['valor'] == u'true'
                traslado.detalletrasladomantenimiento_set.filter(
                    Q(usuario_modificacion=request.user) | Q(usuario_modificacion=None)).update(seleccionado=estado,
                                                                                                usuario_modificacion=request.user,
                                                                                                fecha_modificacion=datetime.now())
                return JsonResponse({"result": "ok", "reload": 'True', 'contador': traslado.contador()})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos"})

        elif action == 'marcartodostraspaso':
            try:
                traspaso = TraspasoActivo.objects.get(pk=int(request.POST['id']))
                estado = request.POST['valor'] == u'true'
                traspaso.detalletraspasoactivo_set.all().update(seleccionado=estado)
                return JsonResponse({"result": "ok", "reload": 'True', 'contador': traspaso.contador_traspaso()})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos"})

        elif action == 'marcartodosbaja':
            try:
                baja = BajaActivo.objects.get(pk=int(request.POST['id']))
                estado = request.POST['valor'] == u'true'
                baja.detallebajaactivo_set.all().update(seleccionado=estado)
                return JsonResponse({"result": "ok", "reload": 'True', 'contador': baja.contador()})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos"})

        elif action == 'marcarseleccionadotraslado':
            try:
                detalle = DetalleTrasladoMantenimiento.objects.get(pk=int(request.POST['id']))
                detalle.seleccionado = request.POST['valor'] == u'true'
                detalle.save(request)
                return JsonResponse({"result": "ok", "reload": 'False', 'contador': detalle.codigotraslado.contador()})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos"})

        elif action == 'marcarseleccionadotraspaso':
            try:
                detalle = DetalleTraspasoActivo.objects.get(pk=int(request.POST['id']))
                detalle.seleccionado = request.POST['valor'] == u'true'
                detalle.save(request)
                traspaso = detalle.codigotraspaso
                traspaso.totalbienes = traspaso.cantidad_seleccionados()
                return JsonResponse(
                    {"result": "ok", "reload": 'False', 'contador': detalle.codigotraspaso.contador_traspaso()})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos"})

        elif action == 'marcarseleccionadobaja':
            try:
                detalle = DetalleBajaActivo.objects.get(pk=int(request.POST['id']))
                detalle.seleccionado = request.POST['valor'] == u'true'
                detalle.save(request)
                return JsonResponse({"result": "ok", "reload": 'False', 'contador': detalle.codigobaja.contador()})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos"})

        elif action == 'actualizaobservaciontraslado':
            try:
                detalle = DetalleTrasladoMantenimiento.objects.get(pk=int(request.POST['id']))
                detalle.observacion = request.POST['valor']
                detalle.save(request)
                return JsonResponse({"result": "ok", "reload": 'False', 'contador': detalle.codigotraslado.contador()})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos"})

        elif action == 'eliminar_otro':
            try:
                detalle = DetalleConstatacionFisica.objects.get(pk=int(request.POST['id']))
                detalle.delete()
                return JsonResponse(
                    {"result": "ok", "reload": 'False', 'contador': detalle.codigoconstatacion.contadores_ni()})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos"})

        elif action == 'eliminar_noiden':
            try:
                detalle = DetalleNoIdentificado.objects.get(pk=int(request.POST['id']))
                detalle.delete()
                return JsonResponse(
                    {"result": "ok", "reload": 'False', 'contador': detalle.codigoconstatacion.contadores_otros()})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos"})

        elif action == 'marcarrequieretrasotro':
            try:
                detalle = DetalleConstatacionFisica.objects.get(pk=int(request.POST['id']))
                detalle.requieretraspaso = request.POST['valrequieretrasotro'] == u'true'
                detalle.save(request)
                return JsonResponse(
                    {"result": "ok", "reload": 'False', 'contador': detalle.codigoconstatacion.contadores_otros()})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos"})

        elif action == 'marcartras':
            try:
                detalle = DetalleConstatacionFisica.objects.get(pk=int(request.POST['id']))
                encontrado = request.POST['valrequieretras'] == u'true'
                if encontrado:
                    detalle.requieretraspaso = True
                else:
                    detalle.requieretraspaso = False
                detalle.save(request)
                return JsonResponse(
                    {"result": "ok", "reload": 'False', 'contador': detalle.codigoconstatacion.contadores_pertenece()})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos"})

        elif action == 'buscarcodigoconstatacion':
            try:
                constatacion = ConstatacionFisica.objects.get(pk=int(request.POST['id']))
                if constatacion.detalleconstatacionfisica_set.filter(
                        Q(activo__codigogobierno=request.POST['codigo']) | Q(
                            activo__codigointerno=request.POST['codigo']), perteneceusuario=True).exists():
                    detalle = constatacion.detalleconstatacionfisica_set.filter(
                        Q(activo__codigogobierno=request.POST['codigo']) | Q(
                            activo__codigointerno=request.POST['codigo']), perteneceusuario=True)[0]
                    detallecons = constatacion.detalleconstatacionfisica_set.filter(perteneceusuario=True).order_by(
                        'activo__codigointerno', 'activo__codigogobierno')
                    pagina = ((list(detallecons.values_list('id', flat=True)).index(detalle.id) + 1) / 10.0)
                    try:
                        rest = 0
                        if len(str(pagina).split('.')) > 1:
                            rest = (1 if int(str(pagina).split('.')[1]) > 0 else 0)
                        pagina = int(str(pagina).split('.')[0]) + rest
                    except Exception as ex:
                        pass
                    return JsonResponse(
                        {"result": "ok", "reload": 'True', 'encontrado': 'True', 'pertenece': 'True', "pagina": pagina,
                         "id": detalle.id})
                else:
                    if ActivoFijo.objects.filter(
                            Q(codigogobierno=request.POST['codigo']) | Q(codigointerno=request.POST['codigo']),
                            statusactivo=1).exists():
                        if constatacion.detalleconstatacionfisica_set.filter(
                                Q(activo__codigogobierno=request.POST['codigo']) | Q(
                                    activo__codigointerno=request.POST['codigo']), perteneceusuario=False).exists():
                            return JsonResponse({"result": "bad", "mensaje": u"Activo ya adicionado"})
                        activo = ActivoFijo.objects.filter(
                            Q(codigogobierno=request.POST['codigo']) | Q(codigointerno=request.POST['codigo']),
                            statusactivo=1).distinct()[0]
                        detalle = DetalleConstatacionFisica(codigoconstatacion=constatacion,
                                                            perteneceusuario=False,
                                                            activo=activo,
                                                            usuariobienes=activo.responsable,
                                                            ubicacionbienes=activo.ubicacion,
                                                            estadooriginal=activo.estado,
                                                            estadoactual=activo.estado,
                                                            encontrado=True)
                        detalle.save(request)
                        detallecons = constatacion.detalleconstatacionfisica_set.filter(
                            perteneceusuario=False).order_by('activo__codigointerno', 'activo__codigogobierno')
                        pagina = ((list(detallecons.values_list('id', flat=True)).index(detalle.id) + 1) / 10.0)
                        try:
                            rest = 0
                            if len(str(pagina).split('.')) > 1:
                                rest = (1 if int(str(pagina).split('.')[1]) > 0 else 0)
                            pagina = int(str(pagina).split('.')[0]) + rest
                        except:
                            pass
                        return JsonResponse(
                            {"result": "ok", "reload": 'True', 'encontrado': 'True', 'pertenece': 'False',
                             "pagina": pagina, "id": detalle.id})
                    else:
                        if ActivoFijo.objects.filter(
                                Q(codigogobierno=request.POST['codigo']) | Q(codigointerno=request.POST['codigo']),
                                statusactivo=2).exists():
                            return JsonResponse({"result": "bad", "mensaje": u"El Activo se encuentra dado de baja"})
                        if constatacion.detallenoidentificado_set.filter(codigobarra=request.POST['codigo']).exists():
                            return JsonResponse(
                                {"result": "bad", "mensaje": u"Ya se encuentra registrado como No Identificado"})
                        return JsonResponse({"result": "ok", 'encontrado': 'False', 'pertenece': 'False'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al buscar los datos"})

        elif action == 'buscarcodigotraslado':
            try:
                traslado = TrasladoMantenimiento.objects.get(pk=int(request.POST['id']))
                if not traslado.detalletrasladomantenimiento_set.filter(
                        Q(activo__codigogobierno=request.POST['codigo']) | Q(
                            activo__codigointerno=request.POST['codigo'])).exists():
                    return JsonResponse({"result": "bad", "mensaje": u"El activo no se encuentra en lista"})
                detalle = traslado.detalletrasladomantenimiento_set.filter(
                    Q(activo__codigogobierno=request.POST['codigo']) | Q(activo__codigointerno=request.POST['codigo']))[
                    0]
                detalletrasl = traslado.detalletrasladomantenimiento_set.all().distinct().order_by(
                    'activo__codigointerno', 'activo__codigogobierno')
                pagina = ((list(detalletrasl.values_list('id', flat=True)).index(detalle.id) + 1) / 10.0)
                try:
                    rest = 0
                    if len(str(pagina).split('.')) > 1:
                        rest = (1 if int(str(pagina).split('.')[1]) > 0 else 0)
                    pagina = int(str(pagina).split('.')[0]) + rest
                except:
                    pass
                return JsonResponse({"result": "ok", "encontrado": True, "pagina": pagina, "id": detalle.id})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al buscar los datos"})

        elif action == 'buscarcodigotraspaso':
            try:
                traspaso = TraspasoActivo.objects.get(pk=int(request.POST['id']))
                if not traspaso.detalletraspasoactivo_set.filter(Q(activo__codigogobierno=request.POST['codigo']) | Q(
                        activo__codigointerno=request.POST['codigo'])).exists():
                    return JsonResponse({"result": "bad", "mensaje": u"El activo no se encuentra en lista"})
                detalle = traspaso.detalletraspasoactivo_set.filter(
                    Q(activo__codigogobierno=request.POST['codigo']) | Q(activo__codigointerno=request.POST['codigo']))[
                    0]
                detalletrasp = traspaso.detalletraspasoactivo_set.all().distinct().order_by('activo__codigointerno',
                                                                                            'activo__codigogobierno')
                pagina = ((list(detalletrasp.values_list('id', flat=True)).index(detalle.id) + 1) / 10.0)
                try:
                    rest = 0
                    if len(str(pagina).split('.')) > 1:
                        rest = (1 if int(str(pagina).split('.')[1]) > 0 else 0)
                    pagina = int(str(pagina).split('.')[0]) + rest
                except:
                    pass
                return JsonResponse({"result": "ok", "encontrado": True, "pagina": pagina, "id": detalle.id})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al buscar los datos"})

        elif action == 'buscarcodigobaja':
            try:
                baja = BajaActivo.objects.get(pk=int(request.POST['id']))
                if not baja.detallebajaactivo_set.filter(Q(activo__codigogobierno=request.POST['codigo']) | Q(
                        activo__codigointerno=request.POST['codigo'])).exists():
                    return JsonResponse({"result": "bad", "mensaje": u"El activo no se encuentra en lista"})
                detalle = baja.detallebajaactivo_set.filter(
                    Q(activo__codigogobierno=request.POST['codigo']) | Q(activo__codigointerno=request.POST['codigo']))[
                    0]
                detalletrasp = baja.detallebajaactivo_set.all().distinct().order_by('activo__codigointerno',
                                                                                    'activo__codigogobierno')
                pagina = round((list(detalletrasp.values_list('id', flat=True)).index(detalle.id) + 1) / 9.0)
                try:
                    rest = 0
                    if str(pagina).split('.') > 1:
                        rest = (1 if int(str(pagina).split('.')[1]) > 0 else 0)
                    pagina = int(str(pagina).split('.')[0]) + rest
                except:
                    pass
                return JsonResponse({"result": "ok", "encontrado": True, "pagina": pagina, "id": detalle.id})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al buscar los datos"})

        elif action == 'actualizaestado':
            try:
                detalle = DetalleConstatacionFisica.objects.get(pk=int(request.POST['id']))
                detalle.estadoactual_id = request.POST['valestado']
                detalle.save(request)
                return JsonResponse(
                    {"result": "ok", "reload": 'False', 'contador': detalle.codigoconstatacion.contadores_pertenece()})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos"})

        elif action == 'actualizaestadootro':
            try:
                detalle = DetalleConstatacionFisica.objects.get(pk=int(request.POST['id']))
                detalle.estadoactual_id = request.POST['valestado']
                detalle.save(request)
                return JsonResponse(
                    {"result": "ok", "reload": 'False', 'contador': detalle.codigoconstatacion.contadores_otros()})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos"})

        elif action == 'actualizaobservacion':
            try:
                detalle = DetalleConstatacionFisica.objects.get(pk=int(request.POST['id']))
                detalle.observacion = request.POST['valobser']
                detalle.save(request)
                return JsonResponse(
                    {"result": "ok", "reload": 'False', 'contador': detalle.codigoconstatacion.contadores_pertenece()})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos"})

        elif action == 'detallenoidentificado':
            try:
                constatacion = ConstatacionFisica.objects.get(pk=int(request.POST['id']))
                codigobarra = request.POST['codigo']
                if codigobarra:
                    if constatacion.detallenoidentificado_set.filter(codigobarra=request.POST['codigo']).exists():
                        return JsonResponse({"result": "bad", "mensaje": u"Activo ya adicionado"})
                detalle = DetalleNoIdentificado(codigoconstatacion=constatacion,
                                                codigobarra=request.POST['codigo'],
                                                catalogobien_id=int(request.POST['catalogo']),
                                                serie=request.POST['serie'],
                                                descripcion=request.POST['descripcion'],
                                                modelo=request.POST['modelo'],
                                                marca=request.POST['marca'],
                                                estado_id=int(request.POST['codestado']))
                detalle.save(request)
                detallecons = constatacion.detallenoidentificado_set.all().order_by('codigobarra')
                pagina = ((list(detallecons.values_list('id', flat=True)).index(detalle.id)) / 10.0)
                try:
                    pagina = int(pagina) + (1 if str(pagina).split('.')[1] != '00' else 0)
                except:
                    pass
                return JsonResponse({"result": "ok", "mensaje": u"Modificado por otro usuario", "reload": 'True',
                                     'contador': detalle.codigoconstatacion.contadores_ni(), 'pagina': pagina})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos"})

        elif action == 'editcons':
            try:
                constatacion = ConstatacionFisica.objects.get(pk=request.POST['id'])
                f = ConstatacionForm(request.POST)
                if f.is_valid():
                    constatacion.observacion = f.cleaned_data['observacion']
                    constatacion.save(request)
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
            return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'edittraspaso':
            try:
                f = TraspasoActivoForm(request.POST)
                if f.is_valid():
                    traspaso = TraspasoActivo.objects.get(pk=request.POST['id'])
                    traspaso.usuariobienrecibe = f.cleaned_data['usuariobienrecibe']
                    traspaso.ubicacionbienrecibe = f.cleaned_data['ubicacionbienrecibe']
                    traspaso.custodiobienrecibe = f.cleaned_data['custodiobienrecibe']
                    traspaso.tiposolicitud = f.cleaned_data['tiposolicitud']
                    traspaso.oficio = f.cleaned_data['oficio']
                    traspaso.fechaoficio = f.cleaned_data['fechaoficio']
                    traspaso.solicitante = f.cleaned_data['solicitante']
                    traspaso.observacion = f.cleaned_data['observacion']
                    traspaso.save(request)
                    return JsonResponse({"result": "ok", "id": traspaso.id})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'edittraspasocustodio':
            try:
                f = TraspasoActivoForm(request.POST)
                if f.is_valid():
                    traspaso = TraspasoActivo.objects.get(pk=request.POST['id'])
                    traspaso.custodiobienrecibe = f.cleaned_data['custodiobienrecibe']
                    traspaso.usuariobienrecibe = f.cleaned_data['usuariobienrecibe']
                    traspaso.tiposolicitud = f.cleaned_data['tiposolicitud']
                    traspaso.oficio = f.cleaned_data['oficio']
                    traspaso.fechaoficio = f.cleaned_data['fechaoficio']
                    traspaso.solicitante = f.cleaned_data['solicitante']
                    traspaso.observacion = f.cleaned_data['observacion']
                    traspaso.save(request)
                    return JsonResponse({"result": "ok", "id": traspaso.id})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'editbaja':
            try:
                f = BajaActivoForm(request.POST)
                if f.is_valid():
                    baja = BajaActivo.objects.get(pk=request.POST['id'])
                    # baja.usuariorecibe = f.cleaned_data['usuariorecibe']
                    baja.oficio = f.cleaned_data['oficio']
                    baja.tipobaja = f.cleaned_data['tipobaja']
                    baja.fechaoficio = f.cleaned_data['fechaoficio']
                    baja.tiposolicitud = f.cleaned_data['tiposolicitud']
                    baja.solicitante = f.cleaned_data['solicitante']
                    baja.usuarioejecuta = f.cleaned_data['usuarioejecuta']
                    if f.cleaned_data['usuariobienentrega']:
                        baja.usuariobienentrega = f.cleaned_data['usuariobienentrega']
                    if f.cleaned_data['custodioentrega']:
                        baja.custodioentrega = f.cleaned_data['custodioentrega']
                    if f.cleaned_data['ubicacionbienentrega']:
                        baja.ubicacionbienentrega = f.cleaned_data['ubicacionbienentrega']

                    baja.observacion = f.cleaned_data['observacion']
                    # baja.experto = f.cleaned_data['experto']
                    # baja.contadorper = f.cleaned_data['contadorper']
                    baja.save(request)
                    return JsonResponse({"result": "ok", "id": baja.id})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'edittipobaja':
            try:
                f = TipoBajaForm(request.POST)
                if f.is_valid():
                    tipobaja = TipoBaja.objects.get(pk=request.POST['id'])
                    tipobaja.nombre = f.cleaned_data['nombre']
                    tipobaja.save(request)
                    log(u'Modifico Tipo Baja: %s' % tipobaja, request, "edit")
                    return JsonResponse({"result": "ok", "id": tipobaja.id})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'editacta':
            try:
                acta = TraspasoActivo.objects.get(pk=request.POST['id'])
                f = ActasForm(request.POST)
                if f.is_valid():
                    acta.observacion = f.cleaned_data['observacion']
                    acta.proveedor = f.cleaned_data['proveedor']
                    acta.custodiobienrecibe = f.cleaned_data['custodio']
                    acta.save(request)
                    for d in acta.detalletraspasoactivo_set.all():
                        activo = ActivoFijo.objects.get(pk=int(d.activo.id))
                        activo.custodio = f.cleaned_data['custodio']
                        activo.tipocomprobante = f.cleaned_data['tipocomprobante']
                        activo.save(request)
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
            return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'importar':
            try:
                form = ImportarArchivoXLSForm(request.POST, request.FILES)
                if form.is_valid():
                    hoy = datetime.now().date()
                    nfile = request.FILES['archivo']
                    nfile._name = generar_nombre("importacion_", nfile._name)
                    archivo = ArchivoActivoFijo(nombre='IMPORTACION DE ACTIVOS',
                                                fecha=datetime.now().date(),
                                                archivo=nfile,
                                                estado=1,
                                                tipobien=form.cleaned_data['tipobien'])
                    archivo.save(request)
                    return JsonResponse({"result": "ok"})
                else:
                    return JsonResponse({"result": "bad", "mensaje": u"El archivo no tiene el formato correcto."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'exportar':
            try:
                f = ExportacionForm(request.POST)
                if f.is_valid():
                    exportacion = ExportacionesActivos(fecha=f.cleaned_data['fecha'],
                                                       clasebien=f.cleaned_data['clasebien'],
                                                       tipobien=f.cleaned_data['tipobien'],
                                                       cuentacontable=f.cleaned_data['cuentacontable'],
                                                       estado=1)
                    exportacion.save(request)
                    cursor = connection.cursor()
                    if exportacion.tipobien:
                        cursor.execute(
                            "INSERT INTO sagest_exportacionesactivos_activos (exportacionesactivos_id, activofijo_id) "
                            "SELECT %s, sagest_activofijo.id from sagest_activofijo LEFT JOIN sagest_catalogobien ON sagest_activofijo.catalogo_id = sagest_catalogobien.id WHERE statusactivo=1 AND sagest_activofijo.status=True  AND clasebien=%s AND sagest_catalogobien.tipobien_id = %s  AND cuentacontable_id=%s AND subidogobierno=False;",
                            [exportacion.id, exportacion.clasebien, exportacion.tipobien.id,
                             exportacion.cuentacontable.id])
                    else:
                        cursor.execute(
                            "INSERT INTO sagest_exportacionesactivos_activos (exportacionesactivos_id, activofijo_id) "
                            "SELECT %s, sagest_activofijo.id from sagest_activofijo LEFT JOIN sagest_catalogobien ON sagest_activofijo.catalogo_id = sagest_catalogobien.id WHERE statusactivo=1 AND sagest_activofijo.status=True  AND clasebien=%s AND cuentacontable_id=%s AND subidogobierno=False;",
                            [exportacion.id, exportacion.clasebien, exportacion.cuentacontable.id])
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
            return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'generarexpo':
            try:
                exportacion = ExportacionesActivos.objects.get(pk=request.POST['id'])
                output_folder = os.path.join(os.path.join(SITE_ROOT, 'media', 'activos'))
                try:
                    os.makedirs(output_folder)
                except Exception as ex:
                    pass
                nombre = 'ACTIVOS_' + elimina_tildes(exportacion.fecha.strftime(
                    '%Y%m%d_%H%M%S') + '_' + exportacion.clase() + '_' + exportacion.cuentacontable.descripcion).replace(
                    ' ', '') + ".csv"
                filename = os.path.join(output_folder, nombre)
                with io.open(filename, 'wt', encoding="ascii", errors="replace") as fichero:
                    linea = 1
                    cantidad_total = exportacion.activos.count()
                    for activo in exportacion.activos.all():
                        if linea == 1:
                            pass
                        # if exportacion.tipobien.id == 1:
                        if exportacion:
                            colornombre = ''
                            if activo.color:
                                colornombre = activo.color.nombre

                            fila = u"2-Bienes Muebles,181,0000,0000,5,BLD,%s,%s,,%s,%s,%s,1,Matriz,15-ACTAS DE ENTREGA-RECEPCION,01-Acta de Entrega-Recepcion,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,,,,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s" % (
                                activo.fechaingreso.strftime("%d/%m/%Y"),
                                activo.activo_origen(),
                                activo.catalogo.descripcion.replace(";", " ").replace(":", " ").replace("'",
                                                                                                        " ").replace(
                                    "-", " ").replace("/", " ").replace("&", " ").replace("+", " ").replace("{",
                                                                                                            " ").replace(
                                    "}", " ").replace("*", " ").replace("=", " ").replace('"', ' ').replace(',',
                                                                                                            ' ').encode(
                                    "ascii", "ignore"),
                                activo.catalogo.identificador,
                                activo.descripcion.replace(";", " ").replace(":", " ").replace("'", " ").replace("-",
                                                                                                                 " ").replace(
                                    "/", " ").replace("&", " ").replace("+", " ").replace("{", " ").replace("}",
                                                                                                            " ").replace(
                                    "*", " ").replace("=", " ").replace('"', ' ').replace(',', ' ').encode("ascii",
                                                                                                           "ignore"),
                                activo.tipodoc(),
                                activo.fechacomprobante.strftime("%d/%m/%Y"),
                                activo.codigointerno,
                                activo.activo_estado(),
                                str(activo.costo),
                                activo.activo_deprecia(),
                                activo.responsable.identificacion(),
                                activo.responsable.nombre_completo().encode("ascii", "ignore"),
                                activo.ubicacion.codigo,
                                activo.serie.replace("'", "").replace(";", " ").replace(":", " ").replace("'",
                                                                                                          " ").replace(
                                    "-", " ").replace("/", " ").replace("&", " ").replace("+", " ").replace("{",
                                                                                                            " ").replace(
                                    "}", " ").replace("*", " ").replace("=", " ").replace('"', ' ').replace(',',
                                                                                                            ' ').encode(
                                    "ascii", "ignore"),
                                activo.modelo.replace("'", "").replace(";", " ").replace(":", " ").replace("'",
                                                                                                           " ").replace(
                                    "-",
                                    " ").replace(
                                    "/", " ").replace("&", " ").replace("+", " ").replace("{", " ").replace("}",
                                                                                                            " ").replace(
                                    "*", " ").replace("=", " ").replace('"', ' ').replace(',', ' ').encode("ascii",
                                                                                                           "ignore"),
                                activo.marca.replace(";", " ").replace(":", " ").replace("'", " ").replace("-",
                                                                                                           " ").replace(
                                    "/",
                                    " ").replace(
                                    "&", " ").replace("+", " ").replace("{", " ").replace("}", " ").replace("*",
                                                                                                            " ").replace(
                                    "=", " ").replace('"', ' ').replace(',', ' ').encode("ascii", "ignore"),
                                activo.cuentacontable.cuenta,
                                str(activo.costo),
                                str(activo.valorresidual),
                                str(activo.valorlibros),
                                str(activo.valordepreciacionacumulada),
                                (activo.fechafindeprec.strftime("%d/%m/%Y") if activo.deprecia else ""),
                                (str(activo.vidautil) if activo.deprecia else "0"),
                                colornombre.replace(";", " ").replace(":", " ").replace("'", " ").replace("-",
                                                                                                          " ").replace(
                                    "/", " ").replace("&", " ").replace("+", " ").replace("{", " ").replace("}",
                                                                                                            " ").replace(
                                    "*", " ").replace("=", " ").replace('"', ' ').replace(',', ' ').encode("ascii",
                                                                                                           "ignore"),
                                activo.material.replace(";", " ").replace(":", " ").replace("'", " ").replace("-",
                                                                                                              " ").replace(
                                    "/", " ").replace("&", " ").replace("+", " ").replace("{", " ").replace("}",
                                                                                                            " ").replace(
                                    "*", " ").replace("=", " ").replace('"', ' ').replace(',', ' ').encode("ascii",
                                                                                                           "ignore"),
                                activo.dimensiones.replace(";", " ").replace(":", " ").replace("'", " ").replace("-",
                                                                                                                 " ").replace(
                                    "/", " ").replace("&", " ").replace("+", " ").replace("{", " ").replace("}",
                                                                                                            " ").replace(
                                    "*", " ").replace("=", " ").replace('"', ' ').replace(',', ' ').encode("ascii",
                                                                                                           "ignore"),
                                activo.observacion.replace(";", " ").replace(":", " ").replace("'", " ").replace("-",
                                                                                                                 " ").replace(
                                    "/", " ").replace("&", " ").replace("+", " ").replace("{", " ").replace("}",
                                                                                                            " ").replace(
                                    "*", " ").replace("=", " ").replace('"', ' ').replace(',', ' ').encode("ascii",
                                                                                                           "ignore"))
                            fila = fila.replace('\n', ' ').replace('\r', ' ')
                            if not linea == cantidad_total:
                                fila += '\n'
                            fichero.write(fila)
                            linea += 1
                fichero.close()
                exportacion.ficherocsv.name = "activos/%s" % nombre
                exportacion.save(request)
                return JsonResponse({'result': 'ok', 'archivo': nombre})
            except Exception as ex:
                transaction.set_rollback(True)
                translator = Translator()
                return JsonResponse({"result": "bad", "mensaje": translator.translate(ex.__str__(), 'es').text})

        elif action == 'generarexpoxls':
            try:
                exportacion = ExportacionesActivos.objects.get(pk=request.POST['id'])
                output_folder = os.path.join(os.path.join(SITE_ROOT, 'media', 'activos'))
                style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
                style_nb = easyxf('font: name Times New Roman, color-index blue, bold on',
                                  num_format_str='#,##0.00')
                style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
                title = easyxf(
                    'font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                style1 = easyxf(num_format_str='D-MMM-YY')
                font_style = XFStyle()
                font_style.font.bold = True
                font_style2 = XFStyle()
                font_style2.font.bold = False
                wb = Workbook(encoding='utf-8')
                ws = wb.add_sheet('exp_xls_post_part')
                nombre = 'ACTIVOS_' + elimina_tildes(exportacion.fecha.strftime(
                    '%Y%m%d_%H%M%S') + '_' + exportacion.clase() + '_' + exportacion.cuentacontable.descripcion).replace(
                    ' ', '') + ".xls"
                filename = os.path.join(output_folder, nombre)
                book = xlwt.Workbook()
                date_format = xlwt.XFStyle()
                date_format.num_format_str = 'dd/mm/yyyy'
                sheet1 = book.add_sheet('datos')
                fila = 0
                for r in exportacion.activos.all():
                    i = 0
                    ws.write(fila, 0, '2-Bienes Muebles', font_style2)
                    ws.write(fila, 1, '181', font_style2)
                    ws.write(fila, 2, '0000', font_style2)
                    ws.write(fila, 3, '0000', font_style2)
                    ws.write(fila, 4, '5', font_style2)
                    ws.write(fila, 5, 'BLD', font_style2)
                    ws.write(fila, 6, r.fechaingreso, style1)
                    ws.write(fila, 7, r.activo_origen(), font_style2)
                    ws.write(fila, 8, ' ', font_style2)
                    ws.write(fila, 9, r.catalogo.descripcion, font_style2)
                    ws.write(fila, 10, r.catalogo.identificador, font_style2)
                    ws.write(fila, 11, r.descripcion, font_style2)
                    ws.write(fila, 12, '1', font_style2)
                    ws.write(fila, 13, 'Matriz', font_style2)
                    ws.write(fila, 14, '15-ACTAS DE ENTREGA-RECEPCION', font_style2)
                    ws.write(fila, 15, '01-Acta de Entrega-Recepcion', font_style2)
                    ws.write(fila, 16, r.tipodoc(), font_style2)
                    ws.write(fila, 17, r.fechacomprobante, style1)
                    ws.write(fila, 18, r.codigointerno, font_style2)
                    ws.write(fila, 19, r.activo_estado(), font_style2)
                    ws.write(fila, 20, str(r.costo), font_style2)
                    ws.write(fila, 21, r.activo_deprecia(), font_style2)
                    ws.write(fila, 22, r.responsable.identificacion(), font_style2)
                    ws.write(fila, 23, r.responsable.nombre_completo(), font_style2)
                    ws.write(fila, 24, r.ubicacion.codigo, font_style2)
                    ws.write(fila, 25, r.serie, font_style2)
                    ws.write(fila, 26, r.modelo, font_style2)
                    ws.write(fila, 27, r.marca, font_style2)
                    ws.write(fila, 28, ' ', font_style2)
                    ws.write(fila, 29, ' ', font_style2)
                    ws.write(fila, 30, ' ', font_style2)
                    ws.write(fila, 31, r.cuentacontable.cuenta, font_style2)
                    ws.write(fila, 32, str(r.costo), font_style2)
                    ws.write(fila, 33, str(r.valorresidual), font_style2)
                    ws.write(fila, 34, str(r.valorlibros), font_style2)
                    ws.write(fila, 35, str(r.valordepreciacionacumulada), font_style2)
                    ws.write(fila, 36, (r.fechafindeprec if r.deprecia else ""), style1)
                    ws.write(fila, 37, (str(r.vidautil) if r.deprecia else "0"), font_style2)
                    colornombre = ''
                    if r.color:
                        colornombre = r.color.nombre

                    ws.write(fila, 38, colornombre, font_style2)
                    ws.write(fila, 39, r.material, font_style2)
                    ws.write(fila, 40, r.dimensiones, font_style2)
                    ws.write(fila, 41, r.observacion, font_style2)
                    fila += 1
                wb.save(filename)
                exportacion.ficheroxls.name = "activos/%s" % nombre
                exportacion.save(request)
                return JsonResponse({'result': 'ok', 'archivo': nombre})
            except Exception as ex:
                transaction.set_rollback(True)
                translator = Translator()
                return JsonResponse({"result": "bad", "mensaje": translator.translate(ex.__str__(), 'es').text})

        elif action == 'reportebajas':
            with transaction.atomic():
                try:
                    form = DescargarBajasAniosForm(request.POST)
                    todos = request.POST.get('todos','')
                    if todos:
                        form.fields['anioinicio'].required = False
                        form.fields['aniofin'].required = False
                    if form.is_valid():
                        data['anioinicio'] = form.cleaned_data['anioinicio']
                        data['aniofin'] = form.cleaned_data['aniofin']
                        data['todos'] = form.cleaned_data['todos']
                        titulo = f'Generación de reporte de bajas en proceso.'
                        noti = Notificacion(cuerpo='Se inició la descarga del reporte, por favor espere.',
                                            titulo=titulo, destinatario=persona,
                                            url='',
                                            prioridad=1, app_label='SGA-SAGEST',
                                            fecha_hora_visible=datetime.now() + timedelta(days=1), tipo=2,
                                            en_proceso=True)
                        noti.save(request)
                        descarga_activos_bajas_background(request=request, data=data, notif=noti.pk).start()
                        messages.success(request, 'Generando reporte')
                        return JsonResponse({"result": False, "mensaje": f'Generando reporte'})
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({"result": True, "mensaje": "Complete los datos requeridos."})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": True, "mensaje": f"{ex}"})

        elif action == 'subir':
            try:
                form = ImportarArchivoXLSForm(request.POST, request.FILES)
                if form.is_valid():
                    importacion = ArchivoActivoFijo.objects.get(pk=int(request.POST['id']))
                    hoy = datetime.now().date()
                    nfile = request.FILES['archivo']
                    nfile._name = generar_nombre("importacion_", nfile._name)
                    importacion.archivo = nfile
                    importacion.save(request)
                    log(u'Modifico archivo de importacion: %s' % importacion, request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                    return JsonResponse({"result": "bad", "mensaje": u"El archivo no tiene el formato correcto."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'informebajapdf':
            mensaje = "Problemas al generar informe de baja."
            try:
                bajaactivo = None
                detalle = None
                director = None

                activofijo = ActivoFijo.objects.get(pk=request.POST['id'])
                if activofijo.informeactivobaja_set.filter(status=True):
                    bajaactivo = activofijo.informeactivobaja_set.filter(status=True)[0]
                    perso = Persona.objects.get(usuario=bajaactivo.usuario_creacion)
                    detalle = bajaactivo.detalleinformeactivobaja_set.filter(status=True)
                    # if DirectorResponsableBaja.objects.filter(actual=True, status=True).exists():
                    #     director = DirectorResponsableBaja.objects.get(actual=True, status=True)
                    #     if not bajaactivo.fecha_creacion.date() > director.fechainicio:
                    #         if DirectorResponsableBaja.objects.filter(fechainicio__lte=bajaactivo.fecha_creacion.date(),
                    #                                                   fechafin__gte=bajaactivo.fecha_creacion.date(),
                    #                                                   status=True).exists():
                    #             director = DirectorResponsableBaja.objects.get(
                    #                 fechainicio__lte=bajaactivo.fecha_creacion.date(),
                    #                 fechafin__gte=bajaactivo.fecha_creacion.date(), status=True)
                if bajaactivo.tipoinforme == 1:
                    dir = 'af_activofijo/informebajapdf.html'
                    director = get_directorresponsablebaja('AT')
                else:
                    # if PersonaDepartamentoFirmas.objects.filter(actualidad=True, status=True).exists():
                    #     director = PersonaDepartamentoFirmas.objects.get(actualidad=True, status=True,
                    #                                                      tipopersonadepartamento_id=1,
                    #                                                      departamentofirma_id=2)
                    #
                    # if PersonaDepartamentoFirmas.objects.filter(status=True,
                    #                                             fechafin__gte=bajaactivo.fecha_creacion.date(),
                    #                                             fechainicio__lte=bajaactivo.fecha_creacion.date(),
                    #                                             tipopersonadepartamento_id=1,
                    #                                             departamentofirma_id=2).exists():
                    #     director = PersonaDepartamentoFirmas.objects.get(status=True,
                    #                                                      fechafin__gte=bajaactivo.fecha_creacion.date(),
                    #                                                      fechainicio__lte=bajaactivo.fecha_creacion.date(),
                    #                                                      tipopersonadepartamento_id=1,
                    #                                                      departamentofirma_id=2)

                    dir = 'af_activofijo/informebajapdfmantenimiento.html'
                    director = get_directorresponsablebaja('AF')
                ruta_hoja_membretada = get_ruta_hoja_membretada()
                return conviert_html_to_pdf(dir,
                                            {'pagesize': 'A4', 'ruta_hoja_membretada': ruta_hoja_membretada,
                                             'informe': activofijo, 'perso': perso, 'director': director,
                                             'bajaactivo': bajaactivo, 'detalle': detalle, 'hoy': datetime.now().date()
                                             })
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos. %s" % ex})

        elif action == 'procesarimp':
            try:
                importacion = ArchivoActivoFijo.objects.get(pk=request.POST['id'])
                workbook = load_workbook(importacion.archivo.file.name)
                workbookincidencias = xlwt.Workbook()
                sheetincidencias = workbookincidencias.add_sheet('Incidencias')
                sheetcorrectos = workbookincidencias.add_sheet('Activosguardados')
                sheet = workbook._sheets[0]
                listaactas = []
                if importacion.tipobien.id == 2:
                    # VERIFICACIÓN
                    # VEHICULOS
                    linea = 1
                    errorfichero = False
                    for rowx in sheet.rows:
                        errores = False
                        if linea >= 3:
                            codigo = str(sheet.cell(row=linea, column=1).value)
                            cedula = str(sheet.cell(row=linea, column=26).value)
                            tipo = str(sheet.cell(row=linea, column=43).value)
                            if not Persona.objects.filter(cedula__icontains=cedula).exists():
                                sheetincidencias.write(linea, 0, codigo)
                                sheetincidencias.write(linea, 1, 'NO EXISTE EL USUARIO')
                                errores = True
                                errorfichero = True
                            elif not CuentaContable.objects.filter(
                                    cuenta=sheet.cell(row=linea, column=31).value).exists():
                                sheetincidencias.write(linea, 0, codigo)
                                sheetincidencias.write(linea, 1, 'NO EXISTE LA CUENTA CONTABLE')
                                errores = True
                                errorfichero = True
                            elif sheet.cell(row=linea, column=42).value is None:
                                sheetincidencias.write(linea, 0, codigo)
                                sheetincidencias.write(linea, 1, 'NO TIENE DESCRIPCION')
                                errores = True
                                errorfichero = True
                            elif sheet.cell(row=linea, column=46).value is None:
                                sheetincidencias.write(linea, 0, codigo)
                                sheetincidencias.write(linea, 1, 'NO TIENE OBSERVACION')
                                errores = True
                                errorfichero = True
                            elif sheet.cell(row=linea, column=44).value is None:
                                sheetincidencias.write(linea, 0, codigo)
                                sheetincidencias.write(linea, 1, 'NO TIENE N° COMPROBANTE')
                                errores = True
                                errorfichero = True
                            elif sheet.cell(row=linea, column=45).value is None:
                                sheetincidencias.write(linea, 0, codigo)
                                sheetincidencias.write(linea, 1, 'NO TIENE FECHA COMPROBANTE')
                                errores = True
                                errorfichero = True
                            elif not TipoDocumento.objects.filter(codigo__icontains=tipo).exists():
                                sheetincidencias.write(linea, 0, codigo)
                                sheetincidencias.write(linea, 1, 'NO TIENE TIPO COMPROBANTE')
                                errores = True
                                errorfichero = True
                        linea += 1
                    nombre = 'incidencias_activos_' + str(importacion.id) + '.xls'
                    nombrearchivo = os.path.join(MEDIA_ROOT, 'activos', nombre)
                    workbookincidencias.save(nombrearchivo)
                    importacion.archivoincidencias.name = "activos/%s" % nombre
                    importacion.save(request)
                    if errorfichero:
                        return JsonResponse(
                            {"result": "ok", "obs": True, "archivo": importacion.archivoincidencias.url})
                    # ingreso
                    linea = 1
                    contadorguardados = 0
                    for rowx in sheet.rows:
                        if linea >= 3:
                            habilitado = sheet.cell(row=linea, column=47).value
                            numeroacta = sheet.cell(row=linea, column=4).value
                            if habilitado.upper() in ('HABILITADO', 'S', 'SI', 'APROBADO') and numeroacta:
                                codigo = str(sheet.cell(row=linea, column=1).value)
                                tipo = TipoDocumento.objects.filter(
                                    codigo__icontains=str(sheet.cell(row=linea, column=43).value))[0]
                                if not ActivoFijo.objects.filter(codigogobierno=codigo).exists():
                                    codigocatalogo = str(sheet.cell(row=linea, column=3).value)
                                    if not CatalogoBien.objects.filter(identificador=codigocatalogo).exists():
                                        catalogo = CatalogoBien(identificador=codigocatalogo,
                                                                descripcion=str(sheet.cell(row=linea, column=6).value),
                                                                tipobien_id=2)
                                        catalogo.save(request)
                                        log(u'Adicionó catalogo del bien: %s' % catalogo, request, "add")
                                    else:
                                        catalogo = CatalogoBien.objects.filter(identificador=codigocatalogo)[0]
                                    cedula = str(sheet.cell(row=linea, column=26).value)
                                    usuario = Persona.objects.filter(cedula__icontains=cedula)[0]
                                    if not Ubicacion.objects.filter(
                                            nombre=sheet.cell(row=linea, column=25).value).exists():
                                        ubicacion = Ubicacion(nombre=sheet.cell(row=linea, column=25).value,
                                                              codigo='',
                                                              observacion='')
                                        ubicacion.save(request)
                                        log(u'Adicionó ubicación: %s' % ubicacion, request, "add")
                                    else:
                                        ubicacion = \
                                            Ubicacion.objects.filter(nombre=sheet.cell(row=linea, column=25).value)[0]
                                    custodio = ubicacion.responsable
                                    if not ClaseVehiculo.objects.filter(
                                            nombre=sheet.cell(row=linea, column=14).value).exists():
                                        clasevehiculo = ClaseVehiculo(nombre=sheet.cell(row=linea, column=14).value)
                                        clasevehiculo.save(request)
                                        log(u'Adicionó clase de vehículo: %s' % clasevehiculo, request, "add")
                                    else:
                                        clasevehiculo = \
                                            ClaseVehiculo.objects.filter(nombre=sheet.cell(row=linea, column=14).value)[
                                                0]
                                    if not TipoVehiculo.objects.filter(
                                            nombre__icontains=sheet.cell(row=linea, column=15).value).exists():
                                        tipovehiculo = TipoVehiculo(nombre=sheet.cell(row=linea, column=15).value)
                                        tipovehiculo.save(request)
                                        log(u'Adicionó tipo de vehículo: %s' % tipovehiculo, request, "add")
                                    else:
                                        tipovehiculo = \
                                            TipoVehiculo.objects.filter(nombre=sheet.cell(row=linea, column=15).value)[
                                                0]
                                    if not EstadoProducto.objects.filter(
                                            nombre=sheet.cell(row=linea, column=22).value).exists():
                                        estado = EstadoProducto(nombre=sheet.cell(row=linea, column=22).value)
                                        estado.save(request)
                                        log(u'Adicionó tipo de vehículo: %s' % tipovehiculo, request, "add")
                                    else:
                                        estado = \
                                            EstadoProducto.objects.filter(
                                                nombre=sheet.cell(row=linea, column=22).value)[0]
                                    if not Color.objects.filter(nombre=sheet.cell(row=linea, column=20).value).exists():
                                        colorprimario = Color(nombre=sheet.cell(row=linea, column=20).value)
                                        colorprimario.save(request)
                                        log(u'Adicionó color: %s' % colorprimario, request, "add")
                                    else:
                                        colorprimario = \
                                            Color.objects.filter(nombre=sheet.cell(row=linea, column=20).value)[0]
                                    if not Color.objects.filter(nombre=sheet.cell(row=linea, column=21).value).exists():
                                        colorsecundario = Color(nombre=sheet.cell(row=linea, column=21).value)
                                        colorsecundario.save(request)
                                        log(u'Adicionó color: %s' % colorprimario, request, "add")
                                    else:
                                        colorsecundario = \
                                            Color.objects.filter(nombre=sheet.cell(row=linea, column=21).value)[0]
                                    if not OrigenIngreso.objects.filter(
                                            nombre=sheet.cell(row=linea, column=28).value).exists():
                                        origeningreso = OrigenIngreso(nombre=sheet.cell(row=linea, column=28).value)
                                        origeningreso.save(request)
                                        log(u'Adicionó color: %s' % origeningreso, request, "add")
                                    else:
                                        origeningreso = \
                                            OrigenIngreso.objects.filter(nombre=sheet.cell(row=linea, column=28).value)[
                                                0]
                                    if not sheet.cell(row=linea, column=45).value is None:
                                        fechacomp = sheet.cell(row=linea, column=45).value.date()
                                    else:
                                        fechacomp = None
                                    cuenta = \
                                        CuentaContable.objects.filter(cuenta=sheet.cell(row=linea, column=31).value)[0]
                                    if TraspasoActivo.objects.filter(tipo=1, usuariobienrecibe=usuario,
                                                                     ubicacionbienrecibe=ubicacion, estado=1,
                                                                     archivoactivofijo=importacion).exists():
                                        acta = TraspasoActivo.objects.filter(tipo=1, usuariobienrecibe=usuario,
                                                                             ubicacionbienrecibe=ubicacion, estado=1,
                                                                             archivoactivofijo=importacion)[0]
                                    else:
                                        acta = TraspasoActivo(tipo=1,
                                                              actaentregagobierno=int(numeroacta),
                                                              fecha=datetime.now(),
                                                              usuariobienrecibe=usuario,
                                                              ubicacionbienrecibe=ubicacion,
                                                              responsablebienes_id=RESPONSABLE_BIENES_ID,
                                                              asistentebodega_id=ASISTENTE_BODEGA_ID,
                                                              estado=1,
                                                              custodiobienrecibe=custodio)
                                        acta.save(request)
                                        log(u'Adicionó un traspaso de activo: %s' % acta, request, "add")
                                        secuencia = secuencia_activos(request)
                                        secuencia.numeroasignacion += 1
                                        secuencia.save(request)
                                        acta.numero = secuencia.numeroasignacion
                                        acta.normativaacta = secuencia.normativaacta
                                        acta.save(request)
                                    activo = ActivoFijo(codigogobierno=codigo,
                                                        catalogo=catalogo,
                                                        fechaingreso=sheet.cell(row=linea, column=33).value.date(),
                                                        serie=sheet.cell(row=linea, column=7).value,
                                                        modelo=sheet.cell(row=linea, column=8).value,
                                                        marca=sheet.cell(row=linea, column=9).value,
                                                        clasevehiculo=clasevehiculo,
                                                        tipovehiculo=tipovehiculo,
                                                        custodio=custodio,
                                                        responsable=usuario,
                                                        descripcion=str(sheet.cell(row=linea, column=42).value),
                                                        observacion=str(sheet.cell(row=linea, column=46).value),
                                                        tipocomprobante=tipo,
                                                        numerocomprobante=str(sheet.cell(row=linea, column=44).value),
                                                        fechacomprobante=fechacomp,
                                                        ubicacion=ubicacion,
                                                        numeromotor=str(sheet.cell(row=linea, column=16).value),
                                                        numerochasis=str(sheet.cell(row=linea, column=17).value),
                                                        aniofabricacion=int(sheet.cell(row=linea, column=18).value),
                                                        placa=str(sheet.cell(row=linea, column=19).value),
                                                        colorprimario=colorprimario,
                                                        colorsecundario=colorsecundario,
                                                        estado=estado,
                                                        origeningreso=origeningreso,
                                                        fechaultimadeprec=sheet.cell(row=linea, column=34).value.date(),
                                                        vidautil=int(sheet.cell(row=linea, column=35).value),
                                                        cuentacontable=cuenta,
                                                        costo=float(sheet.cell(row=linea, column=37).value),
                                                        valorresidual=float(sheet.cell(row=linea, column=38).value),
                                                        valorlibros=float(sheet.cell(row=linea, column=39).value),
                                                        statusactivo=1,
                                                        valordepreciacionacumulada=float(
                                                            sheet.cell(row=linea, column=40).value))
                                    activo.save(request)

                                    if catalogo.equipoelectronico:
                                        ultimocodigotics = ActivoTecnologico.objects.filter(codigotic__isnull=False).last()
                                        codigotics_generado = int(ultimocodigotics.codigotic) + 1
                                        activonew = ActivoTecnologico(activotecnologico=activo,
                                                                      codigotic=codigotics_generado)
                                        activonew.save(request)

                                    log(u'Adicionó activo: %s' % activo, request, "add")
                                    detalle = DetalleTraspasoActivo(codigotraspaso=acta, activo=activo)
                                    detalle.save(request)
                                    importacion.actas.add(acta)
                                    contadorguardados += 1
                                    for acta in importacion.actas.all():
                                        listaactas.append(acta.numero)
                                    crear_constataciones(activo, persona, request)
                        linea += 1
                elif importacion.tipobien.id == 6:
                    # libros
                    # verificación
                    linea = 1
                    errorfichero = False
                    for rowx in sheet.rows:
                        errores = False
                        if linea >= 3:
                            codigo = str(sheet.cell(row=linea, column=1).value)
                            cedula = str(sheet.cell(row=linea, column=24).value)
                            tipo = str(sheet.cell(row=linea, column=42).value)
                            if not Persona.objects.filter(cedula__icontains=cedula).exists():
                                sheetincidencias.write(linea, 0, codigo)
                                sheetincidencias.write(linea, 1, 'NO EXISTE EL USUARIO')
                                errores = True
                                errorfichero = True
                            elif not CuentaContable.objects.filter(
                                    cuenta=sheet.cell(row=linea, column=29).value).exists():
                                sheetincidencias.write(linea, 0, codigo)
                                sheetincidencias.write(linea, 1, 'NO EXISTE LA CUENTA CONTABLE')
                                errores = True
                                errorfichero = True
                            elif len(remover_caracteres_especiales_unicode(
                                    sheet.cell(row=linea, column=40).value).strip()) == 0:
                                sheetincidencias.write(linea, 0, codigo)
                                sheetincidencias.write(linea, 1, 'NO TIENE DESCRIPCION')
                                errores = True
                                errorfichero = True
                            elif len(remover_caracteres_especiales_unicode(
                                    sheet.cell(row=linea, column=41).value).strip()) == 0:
                                sheetincidencias.write(linea, 0, codigo)
                                sheetincidencias.write(linea, 1, 'NO TIENE OBSERVACION')
                                errores = True
                                errorfichero = True
                            elif len(str(sheet.cell(row=linea, column=43).value).strip()) == 0:
                                sheetincidencias.write(linea, 0, codigo)
                                sheetincidencias.write(linea, 1, 'NO TIENE N° COMPROBANTE')
                                errores = True
                                errorfichero = True
                            elif not Ubicacion.objects.filter(nombre=sheet.cell(row=linea, column=23).value,
                                                              responsable__isnull=False).exists():
                                sheetincidencias.write(linea, 0, codigo)
                                sheetincidencias.write(linea, 1, sheet.cell(row=linea, column=6).value)
                                sheetincidencias.write(linea, 2,
                                                       'CODIGO BIEN SIN UBICACION O RESPONSABLE, VERIFICAR QUE EXISTA LA UBIACION CON RESPONSABLE')
                                errores = True
                                errorfichero = True
                            elif len(str(sheet.cell(row=linea, column=44).value).strip()) == 0:
                                sheetincidencias.write(linea, 0, codigo)
                                sheetincidencias.write(linea, 1, 'NO TIENE FECHA COMPROBANTE')
                                errores = True
                                errorfichero = True
                            elif not TipoDocumento.objects.filter(codigo__icontains=tipo).exists():
                                sheetincidencias.write(linea, 0, codigo)
                                sheetincidencias.write(linea, 1, 'NO TIENE TIPO COMPROBANTE')
                                errores = True
                                errorfichero = True
                        linea += 1
                    nombre = 'incidencias_activos_' + str(importacion.id) + '.xls'
                    nombrearchivo = os.path.join(MEDIA_ROOT, 'activos', nombre)
                    workbookincidencias.save(nombrearchivo)
                    importacion.archivoincidencias.name = "activos/%s" % nombre
                    importacion.save(request)
                    if errorfichero:
                        return JsonResponse(
                            {"result": "ok", "obs": True, "archivo": importacion.archivoincidencias.url})
                    # ingreso
                    linea = 1
                    contadorguardados = 0
                    for rowx in sheet.rows:
                        # cols = sheet.row_values(rowx)
                        if linea >= 3:
                            habilitado = sheet.cell(row=linea, column=21).value
                            numeroacta = sheet.cell(row=linea, column=4).value

                            if habilitado.upper() in ('HABILITADO', 'S', 'SI', 'APROBADO') and numeroacta:
                                codigo = str(sheet.cell(row=linea, column=1).value)
                                tipo = TipoDocumento.objects.filter(
                                    codigo__icontains=str(sheet.cell(row=linea, column=42).value))[0]
                                if not ActivoFijo.objects.filter(codigogobierno=codigo).exists():
                                    codigocatalogo = str(sheet.cell(row=linea, column=3).value)
                                    if not CatalogoBien.objects.filter(identificador=codigocatalogo).exists():
                                        catalogo = CatalogoBien(identificador=codigocatalogo,
                                                                descripcion=str(sheet.cell(row=linea, column=6).value),
                                                                tipobien_id=6)
                                        catalogo.save(request)
                                    else:
                                        catalogo = CatalogoBien.objects.filter(identificador=codigocatalogo)[0]
                                    cedula = str(sheet.cell(row=linea, column=24).value)
                                    usuario = Persona.objects.filter(cedula__icontains=cedula)[0]
                                    if not Ubicacion.objects.filter(
                                            nombre=sheet.cell(row=linea, column=23).value).exists():
                                        ubicacion = Ubicacion(nombre=sheet.cell(row=linea, column=23).value,
                                                              codigo='',
                                                              observacion='')
                                        ubicacion.save(request)
                                    else:
                                        ubicacion = \
                                            Ubicacion.objects.filter(nombre=sheet.cell(row=linea, column=23).value)[0]
                                    custodio = ubicacion.responsable
                                    if not EstadoProducto.objects.filter(
                                            nombre=sheet.cell(row=linea, column=20).value).exists():
                                        estado = EstadoProducto(nombre=sheet.cell(row=linea, column=20).value)
                                        estado.save(request)
                                    else:
                                        estado = \
                                            EstadoProducto.objects.filter(
                                                nombre=sheet.cell(row=linea, column=20).value)[0]
                                    if not OrigenIngreso.objects.filter(
                                            nombre=sheet.cell(row=linea, column=26).value).exists():
                                        origeningreso = OrigenIngreso(nombre=sheet.cell(row=linea, column=26).value)
                                        origeningreso.save(request)
                                    else:
                                        origeningreso = \
                                            OrigenIngreso.objects.filter(nombre=sheet.cell(row=linea, column=26).value)[
                                                0]
                                    if not sheet.cell(row=linea, column=44).value is None:
                                        fechacomp = sheet.cell(row=linea, column=44).value.date()
                                    else:
                                        fechacomp = None
                                    cuenta = \
                                        CuentaContable.objects.filter(cuenta=sheet.cell(row=linea, column=29).value)[0]
                                    ultimadeprec = None
                                    if sheet.cell(row=linea, column=32).value:
                                        ultimadeprec = sheet.cell(row=linea, column=32).value.date()
                                    if TraspasoActivo.objects.filter(tipo=1, usuariobienrecibe=usuario,
                                                                     ubicacionbienrecibe=ubicacion, estado=1,
                                                                     archivoactivofijo=importacion).exists():
                                        acta = TraspasoActivo.objects.filter(tipo=1, usuariobienrecibe=usuario,
                                                                             ubicacionbienrecibe=ubicacion, estado=1,
                                                                             archivoactivofijo=importacion)[0]
                                    else:
                                        acta = TraspasoActivo(tipo=1,
                                                              actaentregagobierno=int(numeroacta),
                                                              fecha=datetime.now(),
                                                              usuariobienrecibe=usuario,
                                                              ubicacionbienrecibe=ubicacion,
                                                              responsablebienes_id=RESPONSABLE_BIENES_ID,
                                                              asistentebodega_id=ASISTENTE_BODEGA_ID,
                                                              estado=1,
                                                              custodiobienrecibe=custodio)
                                        acta.save(request)
                                        secuencia = secuencia_activos(request)
                                        secuencia.numeroasignacion += 1
                                        secuencia.save(request)
                                        acta.numero = secuencia.numeroasignacion
                                        acta.normativaacta = secuencia.normativaacta
                                        acta.save(request)

                                    activo = ActivoFijo(codigogobierno=codigo,
                                                        catalogo=catalogo,
                                                        fechaingreso=sheet.cell(row=linea, column=31).value.date(),
                                                        custodio=custodio,
                                                        responsable=usuario,
                                                        ubicacion=ubicacion,
                                                        tipocomprobante=tipo,
                                                        descripcion=remover_caracteres_especiales_unicode(
                                                            sheet.cell(row=linea, column=40).value),
                                                        observacion=remover_caracteres_especiales_unicode(
                                                            sheet.cell(row=linea, column=41).value),
                                                        numerocomprobante=str(sheet.cell(row=linea, column=43).value),
                                                        fechacomprobante=fechacomp,
                                                        serie=str(sheet.cell(row=linea, column=7).value),
                                                        modelo=str(sheet.cell(row=linea, column=8).value),
                                                        marca=str(sheet.cell(row=linea, column=9).value),
                                                        titulo=remover_caracteres_especiales_unicode(
                                                            sheet.cell(row=linea, column=14).value),
                                                        autor=sheet.cell(row=linea, column=15).value,
                                                        editorial=sheet.cell(row=linea, column=16).value,
                                                        fechaedicion=sheet.cell(row=linea, column=17).value.date(),
                                                        numeroedicion=str(sheet.cell(row=linea, column=18).value),
                                                        clasificacionbibliografica=str(
                                                            sheet.cell(row=linea, column=19).value),
                                                        estado=estado,
                                                        origeningreso=origeningreso,
                                                        fechaultimadeprec=ultimadeprec,
                                                        vidautil=int(sheet.cell(row=linea, column=33).value),
                                                        cuentacontable=cuenta,
                                                        costo=float(sheet.cell(row=linea, column=35).value),
                                                        valorresidual=float(sheet.cell(row=linea, column=36).value),
                                                        valorlibros=float(sheet.cell(row=linea, column=37).value),
                                                        statusactivo=1,
                                                        valordepreciacionacumulada=float(
                                                            sheet.cell(row=linea, column=38).value),
                                                        deprecia=False)
                                    activo.save(request)
                                    detalle = DetalleTraspasoActivo(codigotraspaso=acta,
                                                                    activo=activo)
                                    detalle.save(request)
                                    importacion.actas.add(acta)
                                    contadorguardados += 1
                                    for acta in importacion.actas.all():
                                        listaactas.append(acta.numero)
                                    crear_constataciones(activo, persona, request)
                        linea += 1
                else:
                    # otros
                    # verificación
                    linea = 1
                    errorfichero = False
                    for rowx in sheet.rows:
                        errores = False
                        if linea >= 3:
                            codigo = str(sheet.cell(row=linea, column=1).value)
                            cedula = str(sheet.cell(row=linea, column=21).value)
                            tipo = str(sheet.cell(row=linea, column=38).value)
                            if not Persona.objects.filter(cedula__icontains=cedula).exists():
                                sheetincidencias.write(linea, 0, codigo)
                                sheetincidencias.write(linea, 1, sheet.cell(row=linea, column=6).value)
                                sheetincidencias.write(linea, 2, 'NO EXISTE EL USUARIO')
                                errores = True
                                errorfichero = True
                            elif not CuentaContable.objects.filter(
                                    cuenta=sheet.cell(row=linea, column=26).value).exists():
                                sheetincidencias.write(linea, 0, codigo)
                                sheetincidencias.write(linea, 1, sheet.cell(row=linea, column=6).value)
                                sheetincidencias.write(linea, 2, 'NO EXISTE LA CUENTA CONTABLE')
                                errores = True
                                errorfichero = True
                            elif len((sheet.cell(row=linea, column=37).value).strip()) == 0:
                                sheetincidencias.write(linea, 0, codigo)
                                sheetincidencias.write(linea, 1, sheet.cell(row=linea, column=6).value)
                                sheetincidencias.write(linea, 2, 'NO TIENE DESCRIPCION')
                                errores = True
                                errorfichero = True
                            elif len((sheet.cell(row=linea, column=41).value).strip()) == 0:
                                sheetincidencias.write(linea, 0, codigo)
                                sheetincidencias.write(linea, 1, sheet.cell(row=linea, column=6).value)
                                sheetincidencias.write(linea, 2, 'NO TIENE OBSERVACION')
                                errores = True
                                errorfichero = True
                            elif len(str(sheet.cell(row=linea, column=39).value).strip()) == 0:
                                sheetincidencias.write(linea, 0, codigo)
                                sheetincidencias.write(linea, 1, sheet.cell(row=linea, column=6).value)
                                sheetincidencias.write(linea, 2, 'NO TIENE N° COMPROBANTE')
                                errores = True
                                errorfichero = True
                            elif len(str(sheet.cell(row=linea, column=40).value).strip()) == 0:
                                sheetincidencias.write(linea, 0, codigo)
                                sheetincidencias.write(linea, 1, sheet.cell(row=linea, column=6).value)
                                sheetincidencias.write(linea, 2, 'NO TIENE FECHA COMPROBANTE')
                                errores = True
                                errorfichero = True
                            elif not Ubicacion.objects.filter(nombre=sheet.cell(row=linea, column=20).value,
                                                              responsable__isnull=False).exists():
                                sheetincidencias.write(linea, 0, codigo)
                                sheetincidencias.write(linea, 1, sheet.cell(row=linea, column=6).value)
                                sheetincidencias.write(linea, 2,
                                                       'CODIGO BIEN SIN UBICACION O RESPONSABLE, VERIFICAR QUE EXISTA LA UBIACION CON RESPONSABLE')
                                errores = True
                                errorfichero = True
                            elif not TipoDocumento.objects.filter(codigo__icontains=tipo).exists():
                                sheetincidencias.write(linea, 0, codigo)
                                sheetincidencias.write(linea, 1, sheet.cell(row=linea, column=6).value)
                                sheetincidencias.write(linea, 2, 'NO TIENE TIPO COMPROBANTE')
                                errores = True
                                errorfichero = True
                        linea += 1
                    nombre = 'incidencias_activos_' + str(importacion.id) + '.xls'
                    nombrearchivo = os.path.join(MEDIA_ROOT, 'activos', nombre)
                    workbookincidencias.save(nombrearchivo)
                    importacion.archivoincidencias.name = "activos/%s" % nombre
                    importacion.save(request)
                    if errorfichero:
                        return JsonResponse(
                            {"result": "ok", "obs": True, "archivo": importacion.archivoincidencias.url})
                    # ingreso
                    linea = 1
                    contadorguardados = 0
                    for rowx in sheet.rows:
                        if linea >= 3:
                            habilitado = sheet.cell(row=linea, column=18).value
                            numeroacta = sheet.cell(row=linea, column=4).value

                            if habilitado.upper() in ('HABILITADO', 'S', 'SI', 'APROBADO') and numeroacta:
                                codigo = str(sheet.cell(row=linea, column=1).value)
                                tipo = TipoDocumento.objects.filter(
                                    codigo__icontains=str(sheet.cell(row=linea, column=38).value))[0]
                                # if not ActivoFijo.objects.filter(codigogobierno=codigo).exists():
                                if Ubicacion.objects.filter(nombre=sheet.cell(row=linea, column=20).value,
                                                            responsable__isnull=False).exists():
                                    codigocatalogo = str(sheet.cell(row=linea, column=3).value)
                                    if not CatalogoBien.objects.filter(identificador=codigocatalogo).exists():
                                        catalogo = CatalogoBien(identificador=codigocatalogo,
                                                                descripcion=str(sheet.cell(row=linea, column=6).value),
                                                                tipobien=importacion.tipobien)
                                        catalogo.save(request)
                                    else:
                                        catalogo = CatalogoBien.objects.filter(identificador=codigocatalogo)[0]
                                    cedula = str(sheet.cell(row=linea, column=21).value)
                                    usuario = Persona.objects.filter(cedula__icontains=cedula)[0]
                                    # if not Ubicacion.objects.filter(nombre=sheet.cell(row=linea, column=20).value).exists():
                                    #     ubicacion = Ubicacion(nombre=sheet.cell(row=linea, column=20).value,
                                    #                           codigo='',
                                    #                           observacion='')
                                    #     ubicacion.save(request)
                                    #     log(u'Adicionó ubicacion: %s' % ubicacion, request, "add")
                                    # else:
                                    #     ubicacion = Ubicacion.objects.filter(nombre=sheet.cell(row=linea, column=20).value)[0]
                                    # custodio = ubicacion.responsable

                                    ubicacion = Ubicacion.objects.filter(nombre=sheet.cell(row=linea, column=20).value)[
                                        0]
                                    custodio = ubicacion.responsable
                                    if not EstadoProducto.objects.filter(
                                            nombre=sheet.cell(row=linea, column=17).value).exists():
                                        estado = EstadoProducto(nombre=sheet.cell(row=linea, column=17).value)
                                        estado.save(request)
                                        log(u'Adicionó estado de producto: %s' % estado, request, "add")
                                    else:
                                        estado = \
                                            EstadoProducto.objects.filter(
                                                nombre=sheet.cell(row=linea, column=17).value)[0]
                                    if not OrigenIngreso.objects.filter(
                                            nombre=sheet.cell(row=linea, column=23).value).exists():
                                        origeningreso = OrigenIngreso(nombre=sheet.cell(row=linea, column=23).value)
                                        origeningreso.save(request)
                                        log(u'Adicionó origen de ingreso: %s' % origeningreso, request, "add")
                                    else:
                                        origeningreso = \
                                            OrigenIngreso.objects.filter(nombre=sheet.cell(row=linea, column=23).value)[
                                                0]
                                    if not Color.objects.filter(nombre=sheet.cell(row=linea, column=14).value).exists():
                                        color = Color(nombre=sheet.cell(row=linea, column=14).value)
                                        color.save(request)
                                        log(u'Adicionó color de activo: %s' % color, request, "add")
                                    else:
                                        color = Color.objects.filter(nombre=sheet.cell(row=linea, column=14).value)[0]
                                    if not sheet.cell(row=linea, column=40).value is None:
                                        fechacomp = sheet.cell(row=linea, column=40).value.date()
                                    else:
                                        fechacomp = None
                                    cuenta = \
                                        CuentaContable.objects.filter(cuenta=sheet.cell(row=linea, column=26).value)[0]
                                    ultimadeprec = None
                                    if sheet.cell(row=linea, column=29).value:
                                        ultimadeprec = sheet.cell(row=linea, column=29).value.date()
                                    if not ActivoFijo.objects.filter(codigogobierno=codigo).exists():
                                        if TraspasoActivo.objects.filter(tipo=1, usuariobienrecibe=usuario,
                                                                         ubicacionbienrecibe=ubicacion, estado=1,
                                                                         archivoactivofijo=importacion).exists():
                                            acta = TraspasoActivo.objects.filter(tipo=1, usuariobienrecibe=usuario,
                                                                                 ubicacionbienrecibe=ubicacion,
                                                                                 estado=1,
                                                                                 archivoactivofijo=importacion)[0]
                                        else:
                                            acta = TraspasoActivo(tipo=1,
                                                                  actaentregagobierno=int(numeroacta),
                                                                  responsablebienes_id=RESPONSABLE_BIENES_ID,
                                                                  asistentebodega_id=ASISTENTE_BODEGA_ID,
                                                                  fecha=datetime.now(),
                                                                  usuariobienrecibe=usuario,
                                                                  ubicacionbienrecibe=ubicacion,
                                                                  estado=1,
                                                                  custodiobienrecibe=custodio)
                                            acta.save(request)
                                            log(u'Adicionó acta de traspaso de activo: %s' % acta, request, "add")
                                            secuencia = secuencia_activos(request)
                                            secuencia.numeroasignacion += 1
                                            secuencia.save(request)
                                            acta.numero = secuencia.numeroasignacion
                                            acta.normativaacta = secuencia.normativaacta
                                            acta.save(request)
                                        dimensiones = 'SIN ESPECIFICAR'
                                        if not str(sheet.cell(row=linea, column=16).value) == 'None':
                                            dimensiones = sheet.cell(row=linea, column=16).value

                                        activo = ActivoFijo(codigogobierno=codigo,
                                                            catalogo=catalogo,
                                                            fechaingreso=sheet.cell(row=linea, column=28).value.date(),
                                                            custodio=custodio,
                                                            responsable=usuario,
                                                            ubicacion=ubicacion,
                                                            descripcion=(sheet.cell(row=linea, column=37).value),
                                                            observacion=(sheet.cell(row=linea, column=41).value),
                                                            tipocomprobante=tipo,
                                                            numerocomprobante=str(
                                                                sheet.cell(row=linea, column=39).value),
                                                            fechacomprobante=fechacomp,
                                                            serie=str(sheet.cell(row=linea, column=7).value),
                                                            modelo=sheet.cell(row=linea, column=8).value,
                                                            marca=sheet.cell(row=linea, column=9).value,
                                                            color=color,
                                                            material=sheet.cell(row=linea, column=15).value,
                                                            dimensiones=dimensiones,
                                                            estado=estado,
                                                            origeningreso=origeningreso,
                                                            fechaultimadeprec=ultimadeprec,
                                                            vidautil=int(sheet.cell(row=linea, column=30).value),
                                                            cuentacontable=cuenta,

                                                            costo=float(
                                                                (sheet.cell(row=linea, column=32).value).replace(',',
                                                                                                                 '')) if type(
                                                                sheet.cell(row=linea,
                                                                           column=32).value) == 'str' else sheet.cell(
                                                                row=linea, column=32).value,
                                                            valorresidual=float(
                                                                (sheet.cell(row=linea, column=33).value).replace(',',
                                                                                                                 '')) if type(
                                                                sheet.cell(row=linea,
                                                                           column=33).value) == 'str' else sheet.cell(
                                                                row=linea, column=33).value,
                                                            valorlibros=float(
                                                                (sheet.cell(row=linea, column=34).value).replace(',',
                                                                                                                 '')) if type(
                                                                sheet.cell(row=linea,
                                                                           column=34).value) == 'str' else sheet.cell(
                                                                row=linea, column=34).value,

                                                            statusactivo=1,

                                                            valordepreciacionacumulada=float(
                                                                (sheet.cell(row=linea, column=35).value).replace(',',
                                                                                                                 '')) if type(
                                                                sheet.cell(row=linea,
                                                                           column=35).value) == 'str' else float(
                                                                sheet.cell(row=linea, column=35).value)
                                                            )

                                        activo.save(request)
                                        log(u'Adicionó activo: %s' % activo, request, "add")
                                        detalle = DetalleTraspasoActivo(codigotraspaso=acta,
                                                                        activo=activo)
                                        detalle.save(request)
                                        importacion.actas.add(acta)
                                        contadorguardados += 1
                                        for acta in importacion.actas.all():
                                            listaactas.append(acta.numero)
                                        crear_constataciones(activo, persona, request)
                                    else:
                                        act = ActivoFijo.objects.get(codigogobierno=codigo, status=True)
                                        act.responsable = usuario
                                        act.custodio = custodio
                                        act.ubicacion = ubicacion
                                        act.save(request)
                                        log(u'Actualizó activo: %s' % act, request, "add")

                        linea += 1
                importacion.estado = 2
                importacion.numimportados = contadorguardados
                importacion.save(request)
                archivoactivo = ArchivoActivoFijo.objects.get(pk=request.POST['id'])
                miarchivoupdate = openpyxl.load_workbook(archivoactivo.archivo.file.name)
                sheetes = miarchivoupdate.active
                maximo_column = (sheet.max_column) + 1
                columnfila = maximo_column - 1
                sheetes.cell(row=1, column=maximo_column).value = maximo_column
                sheetes.cell(row=2, column=maximo_column).value = 'accion'

                columcustodio = maximo_column + 1
                columnfilacustodio = columcustodio - 1
                sheetes.cell(row=1, column=columcustodio).value = columcustodio
                sheetes.cell(row=2, column=columcustodio).value = 'custodio'
                miarchivoupdate.save(archivoactivo.archivo.file.name)

                miarchivo = openpyxl.load_workbook(archivoactivo.archivo.file.name)
                hojas = miarchivo.get_sheet_names()
                listadoar = miarchivo.get_sheet_by_name(str(hojas[0]))
                totallista = list(listadoar.rows)
                a = 0
                for filas in totallista[:]:
                    filas
                    a += 1
                    if a > 2:
                        if importacion.tipobien.id == 6:
                            if ActivoFijo.objects.filter(codigogobierno=filas[0].value, status=True).exists():
                                filas[columnfila].value = 'CORRECTO'
                                ubicacion = Ubicacion.objects.filter(nombre=sheet.cell(row=a, column=23).value)[0]
                                custodio = ubicacion.responsable
                                filas[
                                    columnfilacustodio].value = custodio.apellido1 + ' ' + custodio.apellido2 + ' ' + custodio.nombres
                            else:
                                filas[columnfila].value = 'ERROR AL SUBIR'
                                ubicacion = Ubicacion.objects.filter(nombre=sheet.cell(row=a, column=23).value)[0]
                                custodio = ubicacion.responsable
                                filas[
                                    columnfilacustodio].value = custodio.apellido1 + ' ' + custodio.apellido2 + ' ' + custodio.nombres
                        elif importacion.tipobien.id == 2:
                            if ActivoFijo.objects.filter(codigogobierno=filas[0].value, status=True).exists():
                                filas[columnfila].value = 'CORRECTO'
                                ubicacion = Ubicacion.objects.filter(nombre=sheet.cell(row=a, column=25).value)[0]
                                custodio = ubicacion.responsable
                                filas[
                                    columnfilacustodio].value = custodio.apellido1 + ' ' + custodio.apellido2 + ' ' + custodio.nombres
                            else:
                                filas[columnfila].value = 'ERROR AL SUBIR'
                                ubicacion = Ubicacion.objects.filter(nombre=sheet.cell(row=a, column=25).value)[0]
                                custodio = ubicacion.responsable
                                filas[
                                    columnfilacustodio].value = custodio.apellido1 + ' ' + custodio.apellido2 + ' ' + custodio.nombres
                        else:
                            if ActivoFijo.objects.filter(codigogobierno=filas[0].value, status=True).exists():
                                filas[columnfila].value = 'CORRECTO'
                                ubicacion = Ubicacion.objects.filter(nombre=sheet.cell(row=a, column=20).value)[0]
                                custodio = ubicacion.responsable
                                filas[
                                    columnfilacustodio].value = custodio.apellido1 + ' ' + custodio.apellido2 + ' ' + custodio.nombres
                            else:
                                filas[columnfila].value = 'ERROR AL SUBIR'
                                ubicacion = Ubicacion.objects.filter(nombre=sheet.cell(row=a, column=20).value)[0]
                                custodio = ubicacion.responsable
                                filas[
                                    columnfilacustodio].value = custodio.apellido1 + ' ' + custodio.apellido2 + ' ' + custodio.nombres

                miarchivo.save(archivoactivo.archivo.file.name)

                log(u'Proceso archivo de importacion: %s' % importacion, request, "edit")
                return JsonResponse({"result": "ok", "obs": False, "actas": listaactas})
            except Exception as ex:
                transaction.set_rollback(True)
                print(ex)
                print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'confirmacion':
            try:
                traslado = TrasladoMantenimiento.objects.get(pk=int(request.POST['id']))
                traslado.tiporegistro = 2
                secuencia = secuencia_activos(request)
                secuencia.numeromantenimientodefinitivo += 1
                secuencia.save(request)
                traslado.numerodefinitivo = secuencia.numeromantenimientodefinitivo
                traslado.fechadefinitivo = datetime.now().date()
                traslado.save(request)
                log(u'Confirmar planificación: %s' % traslado, request, "edit")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'depreciar':
            try:
                fecha = convertir_fecha_invertida(request.POST['fecha'])
                if ActivoFijo.objects.filter(fechaultimadeprec__gt=fecha, statusactivo=1, deprecia=True).exists():
                    activo = ActivoFijo.objects.filter(fechaultimadeprec__gt=fecha, statusactivo=1, deprecia=True)[0]
                    return JsonResponse({"result": "bad",
                                         "mensaje": u"Error la fecha de depreciación es incorrecta: La última depreciación tiene fecha %s" % activo.fechaultimadeprec.strftime(
                                             '%d-%m-%Y')})
                if fecha > datetime.now().date():
                    return JsonResponse({"result": "bad", "mensaje": u"La fecha establecida es mayor a la actual"})
                activo = ActivoFijo.objects.filter(statusactivo=1, deprecia=True).exclude(vidautil=0)
                activos = []
                for ac in activo:
                    act = {'descripcion': unicode(ac.descripcion if len(ac.descripcion) > 0 else "None"), 'id': ac.id}
                    activos.append(act)
                return JsonResponse({"result": "ok", "cantidad": len(activos), "activos": activos})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al generar lista para depreciar"})

        elif action == 'depreciando':
            try:
                fecha = convertir_fecha_invertida(request.POST['fecha'])
                activo = ActivoFijo.objects.get(pk=request.POST['maid'])
                activo.depreciar_activo(fecha, request)
                activo.save(request)
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad"})

        elif action == 'depreciaredificio':
            try:
                fecha = convertir_fecha(request.POST['fecha'])
                if Edificio.objects.filter(fechaultimadeprec__gt=fecha, estado=True, depreciable=True).exists():
                    edificio = Edificio.objects.filter(fechaultimadeprec__gt=fecha, estado=True, depreciable=True)[0]
                    return JsonResponse({"result": "bad",
                                         "mensaje": u"Error la fecha de depreciación es incorrecta: La última depreciación tiene fecha %s" % edificio.fechaultimadeprec.strftime(
                                             '%d-%m-%Y')})
                if fecha > datetime.now().date():
                    return JsonResponse({"result": "bad", "mensaje": u"La fecha establecida es mayor a la actual"})
                edificio = Edificio.objects.filter(estado=True, depreciable=True).exclude(vidautil=0)
                edificios = []
                for ac in edificio:
                    act = {'descripcion': unicode(
                        ac.identificacion.descripcion if len(ac.identificacion.descripcion) > 0 else "None"),
                        'id': ac.id}
                    edificios.append(act)
                return JsonResponse({"result": "ok", "cantidad": len(edificios), "edificios": edificios})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al generar lista para depreciar"})

        elif action == 'depreciandoedificio':
            try:
                fecha = convertir_fecha(request.POST['fecha'])
                edificio = Edificio.objects.get(pk=request.POST['id'])
                edificio.depreciar_edificio(fecha, request)
                edificio.save(request)
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad"})

        elif action == 'manten':
            try:
                detalle = int(request.POST['id'])
                mantenimiento = DetalleTrasladoMantenimiento.objects.get(pk=detalle)
                nombre = mantenimiento.observacion
                return JsonResponse({"result": "ok", "nombre": nombre})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        # elif action == 'detalle_activo':
        #     try:
        #         data['activo'] = activo = ActivoFijo.objects.get(pk=int(request.POST['id']))
        #         template = get_template("af_activofijo/detalle.html")
        #         json_content = template.render(data)
        #         return JsonResponse({"result": "ok", 'html': json_content})
        #     except Exception as ex:
        #         return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'detalle_activohist':
            try:
                data['activo'] = activo = ActivoFijo.objects.get(id=int(request.POST['id']))
                template = get_template("af_activofijo/detalle.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'html': json_content})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'detalle_mantenimiento':
            try:
                data['activofijo'] = activofijo = ActivoFijo.objects.get(pk=int(request.POST['id']))
                data['detallemantenimiento'] = HdDetalle_Incidente.objects.filter(incidente__activo=activofijo,
                                                                                  status=True, incidente__status=True,
                                                                                  estado__id=3).order_by(
                    'incidente__fechareporte')
                sgarantia = activofijo.mantenimientosactivospreventivos_set.filter(status=True).order_by()
                data['mantenimientopreventivo'] = sgarantia.order_by('fecha')
                mantenimientogarantia = activofijo.mantenimientosactivosgarantia_set.filter(status=True)
                data['mantenimientogarantia'] = mantenimientogarantia.order_by('fechainicio')
                data['costomantenimientogarantia'] = \
                    mantenimientogarantia.filter(status=True).aggregate(cantidad=Sum('valor'))['cantidad']
                template = get_template("af_activofijo/detallehelpdesk.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'html': json_content})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'detalle_control':
            try:
                tarjeta = TarjetaControl.objects.get(pk=int(request.POST['id']))
                data['mantenimiento'] = tarjeta.detallemantenimiento_set.all().order_by('-fecha_creacion')[0]
                template = get_template("af_activofijo/detalle_control.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'html': json_content})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'detalle_constatacion':
            try:
                data['constatacion'] = constatacion = ConstatacionFisica.objects.get(pk=int(request.POST['id']))
                data['detallepert'] = detalles = constatacion.detalleconstatacionfisica_set.filter(
                    perteneceusuario=True)
                data['detallenopert'] = detalles = constatacion.detalleconstatacionfisica_set.filter(
                    perteneceusuario=False)
                data['detallenoiden'] = detalles = constatacion.detallenoidentificado_set.all()
                template = get_template("af_activofijo/detallecons.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'html': json_content})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'detalle_bajas':
            try:
                data['baja'] = baja = BajaActivo.objects.get(pk=int(request.POST['id']))
                data['detalles'] = baja.detallebajaactivo_set.all()
                template = get_template("af_activofijo/detallebaja.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'html': json_content})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'detalle_movimiento':
            try:
                data['traspaso'] = traspaso = TraspasoActivo.objects.get(pk=int(request.POST['id']))
                data['detalles'] = traspaso.detalletraspasoactivo_set.filter(seleccionado=True)
                template = get_template("af_activofijo/detalletras.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'html': json_content})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'detalle_actas':
            try:
                data['traspaso'] = traspaso = TraspasoActivo.objects.get(pk=int(request.POST['id']))
                data['detalles'] = traspaso.detalletraspasoactivo_set.filter(codigotraspaso__tipo=1)
                template = get_template("af_activofijo/detalleactas.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'html': json_content})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'detalle_traslados':
            try:
                data['traslado'] = traslado = TrasladoMantenimiento.objects.get(pk=int(request.POST['id']))
                data['detalles'] = traslado.detalletrasladomantenimiento_set.filter(seleccionado=True)
                template = get_template("af_activofijo/detallemantenimiento.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'html': json_content})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'usuariosubicacion':
            try:
                ubicacion = Ubicacion.objects.get(pk=int(request.POST['id']))
                lista = []
                for usuario in Persona.objects.filter(responsableactivo__ubicacion=ubicacion,
                                                      responsableactivo__statusactivo=1).distinct():
                    if [usuario.id, usuario.nombre_completo_inverso()] not in lista:
                        lista.append([usuario.id, usuario.nombre_completo_inverso()])
                return JsonResponse({"result": "ok", "lista": lista})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'ubicacionusuario':
            try:
                usuario = Persona.objects.get(pk=int(request.POST['id']))
                lista = []
                for ubicacion in Ubicacion.objects.filter(activofijo__responsable=usuario,
                                                          activofijo__statusactivo=1).distinct():
                    if [ubicacion.id, ubicacion.nombre] not in lista:
                        lista.append([ubicacion.id, ubicacion.nombre])
                return JsonResponse({"result": "ok", "lista": lista})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'usuariocustodio':
            try:
                responsable = Persona.objects.get(pk=int(request.POST['id']))
                ubicacion = Ubicacion.objects.get(pk=int(request.POST['idp']))
                lista = []
                for custodio in Persona.objects.filter(custodioactivo__ubicacion=ubicacion,
                                                       custodioactivo__responsable=responsable,
                                                       custodioactivo__statusactivo=1).distinct():
                    if [custodio.id, custodio.nombre_completo_inverso()] not in lista:
                        lista.append([custodio.id, custodio.nombre_completo_inverso()])
                return JsonResponse({"result": "ok", "lista": lista})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'custodioentrega':
            try:
                custodio = Persona.objects.get(pk=int(request.POST['id']))
                lista = []
                for ubicacion in Ubicacion.objects.filter(activofijo__custodio=custodio,
                                                          activofijo__statusactivo=1).distinct():
                    if [ubicacion.id, ubicacion.nombre] not in lista:
                        lista.append([ubicacion.id, ubicacion.nombre])
                return JsonResponse({"result": "ok", "lista": lista})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'consult_cat':
            try:
                catalogo = CatalogoBien.objects.get(pk=int(request.POST['id']))
                lista = []
                for responsable in Persona.objects.filter(responsableactivo__catalogo=catalogo).distinct():
                    if [responsable.id, responsable.nombre_completo_inverso()] not in lista:
                        lista.append([responsable.id, responsable.nombre_completo_inverso()])
                return JsonResponse({"result": "ok", "lista": lista})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'activosusuariocustodio':
            try:
                custodio = Persona.objects.get(pk=int(request.POST['id']))
                ubicacion = Ubicacion.objects.get(pk=int(request.POST['idp']))
                responsable = Persona.objects.get(pk=int(request.POST['idu']))
                data = {}
                data['activos'] = responsable.mis_activos_ubicacion_custodio(ubicacion, custodio, responsable)
                template = get_template("af_activofijo/activospersonaubicacion.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'data': json_content})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'activosusuariocustodiotraspaso':
            try:
                custodio = Persona.objects.get(pk=int(request.POST['id']))
                ubicacion = Ubicacion.objects.get(pk=int(request.POST['idp']))
                responsable = Persona.objects.get(pk=int(request.POST['idu']))
                data = {}
                data['activos'] = responsable.mis_activos_ubicacion_custodio(ubicacion, custodio, responsable)
                template = get_template("af_activofijo/activosusuariocustodiotraspaso.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'data': json_content})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'activos_catalogo':
            try:
                responsable = Persona.objects.get(pk=int(request.POST['id']))
                fechai = convertir_fecha(request.POST['fi'])
                fechaf = convertir_fecha(request.POST['ff'])
                catalogo = CatalogoBien.objects.get(pk=int(request.POST['idc']))
                data = {}
                data['activos'] = ActivoFijo.objects.filter(responsable=responsable, catalogo=catalogo,
                                                            fechaingreso__range=(fechai, fechaf), statusactivo=1)
                data['reporte_0'] = obtener_reporte('consulta_activos_catalogo')
                template = get_template("af_activofijo/activoscatalogo.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'data': json_content})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'activos_ubicacion':
            try:
                ubicacion = Ubicacion.objects.get(pk=int(request.POST['id']))
                responsable = Persona.objects.get(pk=int(request.POST['idr']))
                data = {}
                data['activos'] = ActivoFijo.objects.filter(responsable=responsable, ubicacion=ubicacion,
                                                            statusactivo=1)
                template = get_template("af_activofijo/activosubicacion.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'data': json_content})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'activosbaja':
            try:
                responsable = Persona.objects.get(pk=int(request.POST['idp']))
                ubicacion = Ubicacion.objects.get(pk=int(request.POST['id']))
                data = {}
                data['activos'] = responsable.mis_activos_ubicacion(ubicacion)
                template = get_template("af_activofijo/activosbaja.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'data': json_content})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'activocaracteristica':
            try:
                catalogo = CatalogoBien.objects.get(pk=int(request.POST['id']))
                return JsonResponse({"result": "ok", "tipo": catalogo.tipobien.id})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'misactivos':
            try:
                responsable = Persona.objects.get(pk=int(request.POST['idp']))
                ubicacion = Ubicacion.objects.get(pk=int(request.POST['id']))
                data = {}
                data['activos'] = responsable.mis_activos_ubicacion(ubicacion)
                template = get_template("af_activofijo/activosconstatacion.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'data': json_content})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'activosdeotros':
            try:
                data = {}
                activo = None
                json_content = ""
                if ActivoFijo.objects.filter(
                        Q(codigogobierno=request.POST['id']) | Q(codigointerno=request.POST['id'])).exists():
                    data['activo'] = activo = ActivoFijo.objects.filter(
                        Q(codigogobierno=request.POST['id']) | Q(codigointerno=request.POST['id'])).distinct()[0]
                    template = get_template("af_activofijo/activosotros.html")
                    json_content = template.render(data)
                return JsonResponse({"result": "ok", "cantidad": 1 if activo else 0, 'data': json_content,
                                     "id": activo.id if activo else 0})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'activoubicacion':
            try:
                responsable = Persona.objects.get(pk=int(request.POST['id']))
                lista = []
                for ubicacion in Ubicacion.objects.filter(activofijo__responsable=responsable,
                                                          activofijo__statusactivo=1).distinct():
                    if [ubicacion.id, ubicacion.nombre] not in lista:
                        lista.append([ubicacion.id, ubicacion.nombre])
                return JsonResponse({"result": "ok", "lista": lista})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'addestadobien':
            try:
                form = EstadoBienForm(request.POST)
                if form.is_valid():
                    estado = EstadoBien(nombre=form.cleaned_data['nombre'],
                                        descripcion=form.cleaned_data['descripcion'])
                    estado.save(request)
                    log(u'Adicionó estado de Bienes Inmuebles de la Intitución: %s' % estado, request, "add")
                    return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.rollback()
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'editestadobien':
            try:
                form = EstadoBienForm(request.POST)
                if form.is_valid():
                    if not EstadoBien.objects.get(pk=int(request.POST['id'])).esta_activo():
                        estado = EstadoBien.objects.get(pk=int(request.POST['id']))
                        estado.nombre = form.cleaned_data['nombre']
                        estado.descripcion = form.cleaned_data['descripcion']
                        estado.save(request)
                        log(u'Editó estado de Bienes Inmuebles de la Intitución: %s' % estado, request, "edit")
                        return JsonResponse({"result": "ok"})
                    else:
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"El Estado de Inmuebles y Bienes esta activo"})
            except Exception as ex:
                transaction.rollback()
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'delestadobien':
            try:
                if not EstadoBien.objects.get(pk=int(request.POST['id'])).esta_activo():
                    estado = EstadoBien.objects.get(pk=int(request.POST['id']))
                    log(u'Eliminó estado de Bienes Inmuebles de la Intitución: %s' % estado, request, "del")
                    estado.delete()
                    return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.rollback()
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'addcondicionbien':
            try:
                form = CondicionBienForm(request.POST)
                if form.is_valid():
                    condicion = CondicionBien(nombre=form.cleaned_data['nombre'],
                                              descripcion=form.cleaned_data['descripcion'])
                    condicion.save(request)
                    log(u'Adicionó Condición de Bienes Inmuebles de la Intitución: %s' % condicion, request, "add")
                    return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.rollback()
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'editcondicionbien':
            try:
                form = CondicionBienForm(request.POST)
                if form.is_valid():
                    if not CondicionBien.objects.get(pk=int(request.POST['id'])).esta_activo():
                        condicion = EstadoBien.objects.get(pk=int(request.POST['id']))
                        condicion.nombre = form.cleaned_data['nombre']
                        condicion.descripcion = form.cleaned_data['descripcion']
                        condicion.save(request)
                        log(u'Editó Condición de Bienes Inmuebles de la Intitución: %s' % condicion, request, "edit")
                        return JsonResponse({"result": "ok"})
                    else:
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"El Estado de Inmuebles y Bienes esta activo"})
            except Exception as ex:
                transaction.rollback()
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'delcondicionbien':
            try:
                if not CondicionBien.objects.get(pk=int(request.POST['id'])).esta_activo():
                    condicion = CondicionBien.objects.get(pk=int(request.POST['id']))
                    log(u'Eliminó Condición de Bienes Inmuebles de la Intitución: %s' % condicion, request, "del")
                    condicion.delete()
                    return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.rollback()
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'addedificio':
            try:
                form = EdificioForm(request.POST)
                if form.is_valid():
                    edificio = Edificio(codigobien=form.cleaned_data['codigobien'],
                                        codigoanterior=form.cleaned_data['codigoanterior'],
                                        identificador=form.cleaned_data['identificador'],
                                        fechaingreso=form.cleaned_data['fechaingreso'],
                                        catalogo_id=request.POST['catalogo'],
                                        identificacion=form.cleaned_data['identificacion'],
                                        caracteristica=form.cleaned_data['caracteristica'],
                                        propietario=form.cleaned_data['propietario'],
                                        critico=form.cleaned_data['critico'],
                                        clavecatastral=form.cleaned_data['clavecatastral'],
                                        numeropredio=form.cleaned_data['numeropredio'],
                                        numeropiso=form.cleaned_data['numeropiso'],
                                        areaconstruccion=form.cleaned_data['areaconstruccion'],
                                        numeroescritura=form.cleaned_data['numeroescritura'],
                                        fechaescritura=form.cleaned_data['fechaescritura'],
                                        notaria=form.cleaned_data['notaria'],
                                        condicionbien_id=request.POST['condicionbien'],
                                        estadobien_id=request.POST['estadobien'],
                                        cuentacontable_id=request.POST['cuentacontable'],
                                        responsable_id=request.POST['responsable'],
                                        custodio_id=request.POST['custodio'],
                                        vidautil=form.cleaned_data['vidautil'],
                                        depreciable=form.cleaned_data['depreciable'],
                                        valorcontable=form.cleaned_data['valorcontable'],
                                        observacion=form.cleaned_data['observacion'])
                    edificio.save(request)
                    log(u'Adicionó Bienes Inmuebles de la Intitución: %s' % edificio, request, "add")
                    return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.rollback()
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'addmantenimientogarantia':
            try:
                form = MantenimientosActivosGarantiaForm(request.POST, request.FILES)
                newfile = None
                if 'archivo' in request.FILES:
                    newfile = request.FILES['archivo']
                    if newfile:
                        newfilesd = newfile._name
                        ext = newfilesd[newfilesd.rfind("."):]
                        if not ext.lower() == '.pdf':
                            return JsonResponse({"result": "bad", "mensaje": u"Error, solo archivos .pdf."})
                        if newfile.size > 2194304:
                            return JsonResponse({"result": "bad", "mensaje": u"Error, archivo mayor a 2 Mb."})
                        if newfile:
                            newfile._name = generar_nombre("archivo_malla_", newfile._name)
                if 'arcusen' in request.FILES:
                    newfile = request.FILES['arcusen']
                    if newfile:
                        newfilesd = newfile._name
                        ext = newfilesd[newfilesd.rfind("."):]
                        if not ext.lower() == '.pdf':
                            return JsonResponse({"result": "bad", "mensaje": u"Error, solo archivos .pdf."})
                        if newfile.size > 2194304:
                            return JsonResponse({"result": "bad", "mensaje": u"Error, archivo mayor a 2 Mb."})
                        if newfile:
                            newfile._name = generar_nombre("archivo_malla_", newfile._name)
                if form.is_valid():
                    limp = request.POST['id_limp']
                    danio = request.POST['id_dani']
                    id_cpu = int(request.POST['cpu'])
                    # listadoactivos = json.loads(request.POST['lista_items4'])
                    if form.cleaned_data['proveedor'] > 0:
                        proveedor_id = form.cleaned_data['proveedor']
                    else:
                        proveedor_id = None
                    mantenimientogarantia = MantenimientosActivosGarantia(proveedor_id=proveedor_id,
                                                                          valor=form.cleaned_data['valor'],
                                                                          numreporte=form.cleaned_data['numreporte'],
                                                                          fechainicio=form.cleaned_data['fechainicio'],
                                                                          #
                                                                          tipoactivo=form.cleaned_data['tipoactivo'],
                                                                          estusu=form.cleaned_data['estusu'],
                                                                          horamax=form.cleaned_data['horamax'],
                                                                          minutomax=form.cleaned_data['minutomax'],
                                                                          estfrec=form.cleaned_data['estfrec'],
                                                                          estfent=form.cleaned_data['estfent'],
                                                                          observacion=form.cleaned_data['observacion'],
                                                                          activofijo_id=id_cpu,
                                                                          estnuevo=True
                                                                          )
                    mantenimientogarantia.save(request)
                    elementos = limp.split(',')
                    for elemento in elementos:
                        detalle = TareasActivosPreventivosGarantiaLimp(mantenimiento=mantenimientogarantia,
                                                                       grupos_id=elemento)
                        detalle.save()
                    elementosd = danio.split(',')
                    for elemento in elementosd:
                        detalle = TareasActivosPreventivosGarantiaErr(mantenimiento=mantenimientogarantia,
                                                                      grupos_id=elemento)
                        detalle.save()
                    if 'archivo' in request.FILES:
                        newfile = request.FILES['archivo']
                        if newfile:
                            newfile._name = generar_nombre("archivo_garantia", newfile._name)
                            mantenimientogarantia.archivo = newfile
                            mantenimientogarantia.save(request)
                    if 'arcusen' in request.FILES:
                        newfile = request.FILES['arcusen']
                        if newfile:
                            newfile._name = generar_nombre("archivo_entusuario", newfile._name)
                            mantenimientogarantia.arcusen = newfile
                            mantenimientogarantia.save(request)
                    # for codactivo in listadoactivos:
                    #     detallegarantia = DetalleMantenimientosActivosGarantia(mantenimientoactivosgarantia=mantenimientogarantia,
                    #                                                            activofijo_id=codactivo)
                    #     detallegarantia.save(request)
                    log(u'Adicionó mantenimiento de garantia: %s' % mantenimientogarantia, request, "add")
                    return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.rollback()
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'editedificio':
            try:
                form = EdificioForm(request.POST)
                if form.is_valid():
                    edificio = Edificio.objects.get(pk=int(request.POST['id']))
                    edificio.codigobien = request.POST['codigobien']
                    edificio.codigoanterior = request.POST['codigoanterior']
                    edificio.identificador = request.POST['identificador']
                    edificio.fechaingreso = form.cleaned_data['fechaingreso']
                    edificio.catalogo_id = request.POST['catalogo']
                    edificio.identificacion_id = request.POST['identificacion']
                    edificio.caracteristica = form.cleaned_data['caracteristica']
                    edificio.propietario = form.cleaned_data['propietario']
                    edificio.critico = form.cleaned_data['critico']
                    edificio.clavecatastral = request.POST['clavecatastral']
                    edificio.numeropredio = str(request.POST['numeropredio'])
                    edificio.numeropiso = request.POST['numeropiso']
                    edificio.areaconstruccion = request.POST['areaconstruccion']
                    edificio.numeroescritura = str(request.POST['numeroescritura'])
                    edificio.fechaescritura = form.cleaned_data['fechaescritura']
                    edificio.notaria = form.cleaned_data['notaria']
                    edificio.condicionbien_id = request.POST['condicionbien']
                    edificio.estadobien_id = request.POST['estadobien']
                    edificio.cuentacontable_id = request.POST['cuentacontable']
                    edificio.responsable_id = request.POST['responsable']
                    edificio.custodio_id = request.POST['custodio']
                    edificio.vidautil = int(request.POST['vidautil'])
                    edificio.depreciable = form.cleaned_data['depreciable']
                    edificio.valorcontable = form.cleaned_data['valorcontable']
                    edificio.observacion = form.cleaned_data['observacion']
                    edificio.save(request)
                    log(u'Editó Bienes Inmuebles de la Intitución: %s' % edificio, request, "edit")
                    return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.rollback()
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'deledificio':
            try:
                edificio = Edificio.objects.get(pk=int(request.POST['id']))
                log(u'Eliminó Bienes Inmuebles de la Intitución: %s  el usuario %s' % (edificio, usuario), request,
                    "del")
                edificio.delete()
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.rollback()
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'deltaremantenimiento':
            try:
                edificio = MantenimientoGruCategoria.objects.get(pk=int(request.POST['id']))
                log(u'Eliminó tarea de mantenimiento: %s  el usuario %s' % (edificio, usuario), request, "del")
                edificio.delete()
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.rollback()
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'addtareamantenimiento':
            try:
                form = TareaMantenimientoFrom(request.POST)
                if form.is_valid():
                    edificio = MantenimientoGruCategoria(grupocategoria=form.cleaned_data['categoria'],
                                                         descripcion=form.cleaned_data['descripcion'])
                    edificio.save(request)
                    log(u'Adicionó tarea de mantenimiento: %s' % edificio, request, "add")
                    return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.rollback()
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'addtareamantenimientosgdanios':
            try:
                form = TareaMantenimientoDaniosForm(request.POST)
                if form.is_valid():
                    edificio = MantenimientoGruDanios(grupocategoria=form.cleaned_data['categoria'],
                                                      descripcion=(form.cleaned_data['descripcion']).upper())
                    edificio.save(request)
                    log(u'Adicionó tarea de mantenimiento sin garantia en danios: %s' % edificio, request, "add")
                    return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.rollback()
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'edittareamantenimientosgdanio':
            try:
                form = TareaMantenimientoFrom(request.POST)
                if form.is_valid():
                    tarea = MantenimientoGruDanios.objects.get(pk=int(request.POST['id']))
                    tarea.grupocategoria = form.cleaned_data['categoria']
                    tarea.descripcion = form.cleaned_data['descripcion'].upper()
                    tarea.save(request)
                    log(u'Aditó tarea de mantenimiento sin garantia en danio: %s' % tarea, request, "edit")
                    return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.rollback()
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'deltaremantenimientosgdanio':
            try:
                edificio = MantenimientoGruDanios.objects.get(pk=int(request.POST['id']))
                log(u'Eliminó tarea de mantenimiento sin garantia en danio: %s  el usuario %s' % (edificio, usuario),
                    request, "del")
                edificio.delete()
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.rollback()
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'act_des_tareasgdanio':
            try:
                if 'id' in request.POST:
                    tarea = MantenimientoGruDanios.objects.get(status=True, pk=int(request.POST['id']))
                    if tarea.activo:
                        tarea.activo = False
                    else:
                        tarea.activo = True
                    tarea.save(request)
                    return JsonResponse({"result": "ok", "valor": 1 if tarea.activo else 0})
                else:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al activar la tarea."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        # Mantenimiento de Limpieza
        elif action == 'addtareamantenimientolimpieza':
            try:
                form = TareaMantenimientoFrom(request.POST)
                if form.is_valid():
                    limpieza = MantenimientoGruCategoriaGarantiaLimp(grupocategoria=form.cleaned_data['categoria'],
                                                                     descripcion=form.cleaned_data[
                                                                         'descripcion'].upper())
                    limpieza.save(request)
                    log(u'Adicionó tarea de mantenimiento de limpieza: %s' % limpieza, request, "add")
                    return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.rollback()
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'deltaremantenimientolimpieza':
            try:
                limpieza = MantenimientoGruCategoriaGarantiaLimp.objects.get(pk=int(request.POST['id']))
                log(u'Eliminó tarea de mantenimiento en limpieza: %s  el usuario %s' % (limpieza, usuario), request,
                    "del")
                limpieza.delete()
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.rollback()
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'act_des_tarea_limp':
            try:
                if 'id' in request.POST:
                    tarea = MantenimientoGruCategoriaGarantiaLimp.objects.get(status=True, pk=int(request.POST['id']))
                    if tarea.activo:
                        tarea.activo = False
                    else:
                        tarea.activo = True
                    tarea.save(request)
                    return JsonResponse({"result": "ok", "valor": 1 if tarea.activo else 0})
                else:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al activar la tarea."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'edittareamantenimientolimpieza':
            try:
                form = TareaMantenimientoFrom(request.POST)
                if form.is_valid():
                    tarea = MantenimientoGruCategoriaGarantiaLimp.objects.get(pk=int(request.POST['id']))
                    tarea.grupocategoria = form.cleaned_data['categoria']
                    tarea.descripcion = form.cleaned_data['descripcion'].upper()
                    tarea.save(request)
                    log(u'Editó tarea de mantenimiento en limpieza: %s' % tarea, request, "edit")
                    return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.rollback()
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'listatipomantenimientolimpieza':
            try:
                listatareas = MantenimientoGruCategoriaGarantiaLimp.objects.filter(
                    grupocategoria_id=request.POST['id_tipo'], status=True, activo=True)
                lista = []
                for lis in listatareas:
                    lista.append([lis.id, lis.descripcion])
                return JsonResponse({'result': 'ok', 'lista': lista})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        # Mantenimientos en daños
        elif action == 'addtareamantenimientodanio':
            try:
                form = TareaMantenimientoFrom(request.POST)
                if form.is_valid():
                    limpieza = MantenimientoGruCategoriaGarantiaErr(grupocategoria=form.cleaned_data['categoria'],
                                                                    descripcion=form.cleaned_data[
                                                                        'descripcion'].upper())
                    limpieza.save(request)
                    log(u'Adicionó tarea de mantenimiento de danio: %s' % limpieza, request, "add")
                    return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.rollback()
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'act_des_tarea_danio':
            try:
                if 'id' in request.POST:
                    tarea = MantenimientoGruCategoriaGarantiaErr.objects.get(status=True, pk=int(request.POST['id']))
                    if tarea.activo:
                        tarea.activo = False
                    else:
                        tarea.activo = True
                    tarea.save(request)
                    return JsonResponse({"result": "ok", "valor": 1 if tarea.activo else 0})
                else:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al activar la tarea."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'tareasmantenimientodanio':
            try:
                danios = MantenimientoGruCategoriaGarantiaErr.objects.get(pk=int(request.POST['id']))
                log(u'Eliminó tarea de mantenimiento en danios: %s  el usuario %s' % (danios, usuario), request, "del")
                danios.delete()
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.rollback()
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'edittaremantenimientodanio':
            try:
                form = TareaMantenimientoFrom(request.POST)
                if form.is_valid():
                    tarea = MantenimientoGruCategoriaGarantiaErr.objects.get(pk=int(request.POST['id']))
                    tarea.grupocategoria = form.cleaned_data['categoria']
                    tarea.descripcion = form.cleaned_data['descripcion'].upper()
                    tarea.save(request)
                    log(u'Editó tarea de mantenimiento en danios: %s' % tarea, request, "edit")
                    return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.rollback()
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'listatipomantenimientodanio':
            try:
                listatareas = MantenimientoGruCategoriaGarantiaErr.objects.filter(
                    grupocategoria_id=request.POST['id_tipo'], status=True, activo=True)
                lista = []
                for lis in listatareas:
                    lista.append([lis.id, lis.descripcion])
                return JsonResponse({'result': 'ok', 'lista': lista})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'delmantenimiento':
            try:
                mantenimiento = MantenimientosActivosPreventivos.objects.get(pk=int(request.POST['id']))
                log(u'Eliminó mantenimiento: %s  el usuario %s' % (mantenimiento, usuario), request, "del")
                mantenimiento.delete()
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.rollback()
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'delmantenimientogarantia':
            try:
                mantenimientogarantia = MantenimientosActivosGarantia.objects.get(pk=int(request.POST['id']))
                log(u'Eliminó mantenimiento garantía: %s  el usuario %s' % (mantenimientogarantia, usuario), request,
                    "del")
                mantenimientogarantia.delete()
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.rollback()
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'validar_conf':
            try:
                mens = ''
                if not CondicionBien.objects.filter(status=True).exists():
                    mens = u"Condición del Bien, "
                if not EstadoBien.objects.filter(status=True).exists():
                    mens += u"Estado del Bien, "
                if CondicionBien.objects.filter(status=True).exists() and EstadoBien.objects.filter(
                        status=True).exists():
                    return JsonResponse({"result": "ok"})
                else:
                    return JsonResponse(
                        {"result": "bad", "mensaje": "No Tiene Registro Ingresado en " + mens[:mens.__len__() - 2]})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'detalle_edificio':
            try:
                data['edificio'] = Edificio.objects.get(pk=int(request.POST['id']))
                template = get_template("af_activofijo/detalleedificio.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'html': json_content})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'detalle_mantenimientopreven':
            try:
                data['mantenimiento'] = mantenimiento = MantenimientosActivosPreventivos.objects.get(
                    pk=request.POST['id'])
                # data['tareas']  = TareasActivosPreventivos.objects.filter(mantenimiento=mantenimiento)
                data['tareasmantenimiento'] = MantenimientoGruCategoria.objects.filter(
                    grupocategoria=mantenimiento.tipoactivo, status=True)
                data['tareasactivo'] = TareasActivosPreventivos.objects.values_list('grupos_id', flat=True).filter(
                    mantenimiento=mantenimiento, status=True)
                data['ppmantenimiento'] = HdPiezaPartes.objects.filter(grupocategoria=mantenimiento.tipoactivo,
                                                                       status=True)
                data['piezaparteactivo'] = PiezaParteActivosPreventivos.objects.filter(mantenimiento=mantenimiento,
                                                                                       status=True)
                data['daniomantenimiento'] = MantenimientoGruDanios.objects.filter(
                    grupocategoria=mantenimiento.tipoactivo, status=True)
                data['danioactivo'] = TareasActivosPreventivosDanios.objects.filter(mantenimiento=mantenimiento,
                                                                                    status=True)
                data['estdan'] = ESTADO_DANIO
                if mantenimiento.nuevo:
                    template = get_template("af_activofijo/detallemantenimientopreventivov2.html")
                else:
                    template = get_template("af_activofijo/detallemantenimientopreventivo.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'html': json_content})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'detalle_incidentetec':
            try:
                data['mantenimiento'] = mantenimiento = HdDetalle_Incidente.objects.get(id=request.POST['id'])
                template = get_template("af_activofijo/detalleincidentetecnologico.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'html': json_content})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'detalle_mantgarantia':
            try:
                data['mantenimiento'] = mantenimiento = MantenimientosActivosGarantia.objects.get(pk=request.POST['id'])
                data['tareaslimpieza'] = MantenimientoGruCategoriaGarantiaLimp.objects.filter(
                    grupocategoria=mantenimiento.tipoactivo, status=True)
                data['tareasdanio'] = MantenimientoGruCategoriaGarantiaErr.objects.filter(
                    grupocategoria=mantenimiento.tipoactivo, status=True)
                data['tareaslimpiezat'] = TareasActivosPreventivosGarantiaLimp.objects.values_list('grupos_id',
                                                                                                   flat=True).filter(
                    mantenimiento=mantenimiento, status=True)
                data['tareasdaniot'] = TareasActivosPreventivosGarantiaErr.objects.values_list('grupos_id',
                                                                                               flat=True).filter(
                    mantenimiento=mantenimiento, status=True)
                template = get_template("af_activofijo/detallemantenimientogarantia.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'html': json_content})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'act_des_tarea':
            try:
                if 'id' in request.POST:
                    tarea = MantenimientoGruCategoria.objects.get(status=True, pk=int(request.POST['id']))
                    if tarea.activo:
                        tarea.activo = False
                    else:
                        tarea.activo = True
                    tarea.save(request)
                    return JsonResponse({"result": "ok", "valor": 1 if tarea.activo else 0})
                else:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al activar la tarea."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'edittareamantenimiento':
            try:
                form = TareaMantenimientoFrom(request.POST)
                if form.is_valid():
                    tarea = MantenimientoGruCategoria.objects.get(pk=int(request.POST['id']))
                    tarea.grupocategoria = form.cleaned_data['categoria']
                    tarea.descripcion = form.cleaned_data['descripcion']
                    tarea.save(request)
                    log(u'Aditó tarea de mantenimiento: %s' % tarea, request, "edit")
                    return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.rollback()
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

        elif action == 'addmanactivofijo':
            try:
                codigoactivofijo = request.POST['codigoactivofijo']
                mangarantiaid = request.POST['mangarantiaid']
                if DetalleMantenimientosActivosGarantia.objects.filter(mantenimientoactivosgarantia_id=mangarantiaid,
                                                                       activofijo_id=codigoactivofijo).exists():
                    return JsonResponse({"result": "bad", "mensaje": u"Activo ya se encuentra registrado."})
                detalle = DetalleMantenimientosActivosGarantia(mantenimientoactivosgarantia_id=mangarantiaid,
                                                               activofijo_id=codigoactivofijo)
                detalle.save(request)
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'delmanactivofijo':
            try:
                detalle = DetalleMantenimientosActivosGarantia.objects.get(pk=int(request.POST['hidencodigodetalle']))
                detalle.delete()
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        if action == 'addActivoSinCod':
            try:
                f = ActivoFijoForm(request.POST)
                if f.is_valid():
                    if f.cleaned_data['responsable'] or f.cleaned_data['ubicacion']:
                        if not f.cleaned_data['responsable'] or not f.cleaned_data['ubicacion']:
                            return JsonResponse(
                                {"result": "bad", "mensaje": u"Debe seleccionar Responsable y Ubicación."})
                    activo = ActivoFijo(
                        nombre=f.cleaned_data['nombre'],
                        descripcion=f.cleaned_data['descripcion'],
                        serie=f.cleaned_data['serie'],
                        modelo=f.cleaned_data['modelo'],
                        marca=f.cleaned_data['marca'],
                        vidautil=f.cleaned_data['vidautil'],
                        estado=f.cleaned_data['estado'])
                    activo.save(request)
                    if f.cleaned_data['responsable']:
                        activo.ubicacion = f.cleaned_data["ubicacion"]
                        activo.responsable_id = f.cleaned_data["responsable"]

                    log(u'Adicionó Activo: %s' % activo, request, "add")
                    cuenta = CuentaContable.objects.get(pk=int(activo.cuentacontable_id))
                    cuenta.activosfijos = True
                    cuenta.save(request)
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'clasificar':
            try:
                accion = request.POST['accion']
                catalogo = CatalogoBien.objects.get(pk=int(request.POST['id']))
                if accion == 'add':
                    catalogo.equipoelectronico = True
                else:
                    catalogo.equipoelectronico = False
                catalogo.clasificado = True
                catalogo.save(request)

                log(u'Clasificó catálogo: %s' % catalogo, request, "add")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'addgrupo':
            try:
                catalogo = CatalogoBien.objects.get(pk=int(request.POST['id']))
                f = GrupoBienForm(request.POST)
                if f.is_valid():
                    catalogo.grupo = f.cleaned_data['grupo']
                    catalogo.save(request)
                    log(u'Adicionó un grupo a catálogo: %s' % catalogo, request, "add")
                    return JsonResponse({"result": False}, safe=False)
                else:
                    raise NameError('Error al enviar los datos')

            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'permitirfirmar':
            try:
                solicitudtraspaso = SolicitudTraspasoActivos.objects.get(id=int(request.POST['id']))
                solicitudtraspaso.puedefirmar = True
                solicitudtraspaso.save(request)
                log(u'Permite firmar el acta: %s' % solicitudtraspaso, request, "act")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u'Error al permitir firma en el acta'})

        elif action == 'denegarfirma':
            try:
                solicitudtraspaso = SolicitudTraspasoActivos.objects.get(id=int(request.POST['id']))
                solicitudtraspaso.puedefirmar = False
                solicitudtraspaso.save(request)
                log(u'No permite firmar el acta: %s' % solicitudtraspaso, request, "act")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u'Error al permitir firma en el acta'})

        elif action == 'firmardocumento':
            try:
                # Parametros
                txtFirmas = json.loads(request.POST['txtFirmas'])
                if not txtFirmas:
                    raise NameError("Debe seleccionar ubicación de la firma")
                x = txtFirmas[-1]
                idtraspaso = int(encrypt(request.POST['id_objeto']))
                responsables = request.POST.getlist('responsables[]')
                firma = request.FILES["firma"]
                passfirma = request.POST['palabraclave']
                url_archivo = (SITE_STORAGE + request.POST["url_archivo"]).replace('\\', '/')
                _name = generar_nombre(f'actatraspasoactivo{request.user.username}_{idtraspaso}_', 'firmada')
                folder = os.path.join(SITE_STORAGE, 'media', 'traspasoactivofirma', '')

                # Firmar y guardar archivo en folder definido.
                firma = firmararchivogenerado(request, passfirma, firma, url_archivo, folder, _name, x["numPage"],
                                              x["x"], x["y"], x["width"], x["height"])
                if firma != True:
                    raise NameError(firma)
                log(u'Firmo Documento: {}'.format(_name), request, "add")

                folder_save = os.path.join('traspasoactivofirma', '').replace('\\', '/')
                url_file_generado = f'{folder_save}{_name}.pdf'
                actualizaracta = TraspasoActivo.objects.get(id=int(idtraspaso))
                actualizaracta.traspasoactivofirma = url_file_generado
                actualizaracta.save(request)
                log(u'Guardo archivo firmado: {}'.format(actualizaracta), request, "act")
                solicitudtraspaso = SolicitudTraspasoActivos.objects.get(traspasoactivofijo_id=int(idtraspaso))
                if persona.id == 1204:
                    solicitudtraspaso.firmaactivofijo = True
                    seguimiento = SeguimientoSolicitudTraspaso(solicitudtraspaso=solicitudtraspaso, estado=14)
                solicitudtraspaso.save(request)
                log(u'Firman participantes: {}'.format(solicitudtraspaso), request, "act")
                seguimiento.save(request)
                log(u'Seguimiento solicitud traspaso activo: {}'.format(seguimiento), request, "add")
                if solicitudtraspaso.firmaresponsable == True and solicitudtraspaso.firmafuturoresponsable == True and solicitudtraspaso.firmaactivofijo == True and solicitudtraspaso.firmacustodioentrega == True and solicitudtraspaso.firmacustodiorecibe == True:
                    solicitudtraspaso.puedefirmar = False
                    solicitudtraspaso.estado = 6
                    solicitudtraspaso.save(request)
                # return JsonResponse({"result": "ok", "mensaje": "Guardado con exito", }, safe=False)
                return HttpResponseRedirect("/af_activofijo?action=movimientos")
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar. %s" % ex.__str__()}, safe=False)

        elif action == 'generaractatraspaso':
            try:
                solicitudtraspaso = SolicitudTraspasoActivos.objects.get(status=True, id=int(request.POST['id']))
                traspasoactivofijo = TraspasoActivo.objects.get(status=True, id=solicitudtraspaso.traspasoactivofijo_id)
                reporte = Reporte.objects.get(id=211)
                cambiaruta = 0
                isQR = False
                codigo = None
                certificado = None
                base_url = request.META['HTTP_HOST']
                if reporte.archivo:
                    base_url = request.META['HTTP_HOST']
                    d = datetime.now()
                    pdfname = reporte.nombre + d.strftime('%Y%m%d_%H%M%S')
                    tipo = 'pdf'
                    paRequest = {
                        'id': traspasoactivofijo.id,
                        'imp_logo': True,
                        'imp_encabezado': True,
                        'imp_fecha': True,
                        'imp_membretada': False,
                        'url_qr': unicode(base_url + "/".join(
                            [MEDIA_URL, 'documentos', 'userreports', elimina_tildes(request.user.username),
                             pdfname + "." + tipo]))
                    }
                    d = run_report_v1(reporte=reporte, tipo=tipo, paRequest=paRequest, request=request)
                    if not d['isSuccess']:
                        raise NameError(d['mensaje'])
                    else:
                        archivoobtenido = None
                        archivogeneradotraspaso = d['data']['reportfile']
                        data['url_archivo'] = rutaarchivoparafirmar = '{}{}'.format(dominio_sistema,
                                                                                    archivogeneradotraspaso)
                        url_archivo = (SITE_STORAGE + archivogeneradotraspaso).replace('\\', '/')
                        url_archivo = (url_archivo).replace('//', '/')
                        _name = generar_nombre(f'archivotraspasoactivo{request.user.username}_{traspasoactivofijo.id}_',
                                               'generado')
                        folder = os.path.join(SITE_STORAGE, 'media', 'traspasoactivofirma', '')
                        if not os.path.exists(folder):
                            os.makedirs(folder)
                        folder_save = os.path.join('traspasoactivofirma', '').replace('\\', '/')
                        url_file_generado = f'{folder_save}{_name}.pdf'
                        ruta_creacion = SITE_STORAGE
                        ruta_creacion = ruta_creacion.replace('\\', '/')
                        shutil.copy(url_archivo, ruta_creacion + '/media/' + url_file_generado)
                        traspasoactivofijo.traspasoactivofirma = url_file_generado
                        traspasoactivofijo.save()
                        solicitudtraspaso.puedefirmar = True
                        solicitudtraspaso.firmaresponsable = False
                        solicitudtraspaso.firmafuturoresponsable = False
                        solicitudtraspaso.firmaactivofijo = False
                        solicitudtraspaso.firmacustodiorecibe = False
                        solicitudtraspaso.firmacustodioentrega = False
                        solicitudtraspaso.estado = 11
                        solicitudtraspaso.save(request)
                        seguimiento = SeguimientoSolicitudTraspaso(solicitudtraspaso=solicitudtraspaso, estado=19)
                        seguimiento.save(request)
                        seguimiento = SeguimientoSolicitudTraspaso(solicitudtraspaso=solicitudtraspaso, estado=11)
                        seguimiento.save(request)
                        log(u'Actualiza solicitud para poder firmar: {}'.format(solicitudtraspaso), request, "act")
                    return JsonResponse(
                        {"result": "ok", "mensaje": u'El acta de traspaso se ha generado correctamente.'})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u'Error al generar acta de traspaso'})

        # CONSTATACION DE ACTIVOS FIJOS
        elif action == 'addperiodo':
            try:
                form = PeriodoConstatacionAFForm(request.POST)
                if not form.is_valid():
                    transaction.set_rollback(True)
                    form_error = [{k: v[0]} for k, v in form.errors.items()]
                    return JsonResponse({'result': True, "form": form_error, "mensaje": "Error en el formulario"})
                periodo = PeriodoConstatacionAF(nombre=form.cleaned_data['nombre'],
                                                  anio=form.cleaned_data['anio'],
                                                  fechainicio=form.cleaned_data['fechainicio'],
                                                  fechafin=form.cleaned_data['fechafin'],
                                                  baselegal=form.cleaned_data['baselegal'],
                                                  activo=form.cleaned_data['activo'],
                                                  detalle=form.cleaned_data['detalle'])
                periodo.save(request)

                if periodo.activo:
                    periodos=PeriodoConstatacionAF.objects.filter(status=True, activo=True).exclude(id=periodo.id)
                    periodos.update(activo=False)
                log(f'Agrego periodo de constatacion de activo fijo: {periodo}', request, 'add')
                return JsonResponse({'result': False, 'mensaje': 'Guardado con éxito'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': str(ex)})

        elif action == 'editperiodo':
            try:
                periodo=PeriodoConstatacionAF.objects.get(id=int(encrypt(request.POST['id'])))
                form = PeriodoConstatacionAFForm(request.POST, instancia=periodo)
                if not form.is_valid():
                    transaction.set_rollback(True)
                    form_error = [{k: v[0]} for k, v in form.errors.items()]
                    return JsonResponse({'result': True, "form": form_error, "mensaje": "Error en el formulario"})
                periodo.nombre=form.cleaned_data['nombre']
                periodo.anio=form.cleaned_data['anio']
                periodo.fechainicio=form.cleaned_data['fechainicio']
                periodo.fechafin=form.cleaned_data['fechafin']
                periodo.detalle=form.cleaned_data['detalle']
                periodo.baselegal=form.cleaned_data['baselegal']
                periodo.activo=form.cleaned_data['activo']
                periodo.save(request)

                if periodo.activo:
                    periodos=PeriodoConstatacionAF.objects.filter(status=True, activo=True).exclude(id=periodo.id)
                    periodos.update(activo=False)
                log(f'Edito periodo de constatación de activo fijo: {periodo}', request, 'edit')
                return JsonResponse({'result': False, 'mensaje': 'Guardado con éxito'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': str(ex)})

        elif action == 'delperiodo':
            try:
                periodo = PeriodoConstatacionAF.objects.get(id=int(encrypt(request.POST['id'])))
                periodo.status = False
                periodo.save(request)
                log(f'Elimino periodo de constatacion de activos fijos: {periodo}', request, 'del')
                res_json = {"error": False}
            except Exception as ex:
                res_json = {'error': True, "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)

        elif action == 'cerrarperiodo':
            try:
                periodo = PeriodoConstatacionAF.objects.get(id=int(encrypt(request.POST['id'])))
                periodo.totalactivos = total = periodo.total_activos()
                periodo.totalconstatados = total_constatados = periodo.total_constatados()
                periodo.totalpendientes = total-total_constatados
                periodo.fechacierre = hoy
                periodo.activo = False
                periodo.cerrado = True
                periodo.save(request)
                log(f'Se cierra el periodo de constatación: {periodo}', request, 'edit')
                res_json = {"result": 'ok'}
            except Exception as ex:
                res_json = {'result': 'bad', "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)

        elif action == 'marcarencontradoactivo':
            with transaction.atomic():
                try:
                    id_a,id_p=encrypt_id(request.POST['id']), encrypt_id(request.POST['idex'])
                    activo=ActivoFijo.objects.get(id=id_a)
                    detalle_c=DetalleConstatacionFisica.objects.filter(status=True,activo_id=id_a,codigoconstatacion__periodo_id=id_p, codigoconstatacion__status=True).first()
                    if detalle_c and not detalle_c.responsable.id == persona.id:
                        raise NameError(f'Activo que intenta constatar ya fue constatado por: {detalle_c.responsable.nombre_completo_minus()}')
                    if not detalle_c:
                        constatacion = ConstatacionFisica.objects.filter(status=True, usuariobienes=activo.responsable, periodo_id=id_p).first()
                        if not constatacion:
                            # secuencia = secuencia_activos(request)
                            # secuencia.numeroconstatacion += 1
                            # secuencia.save(request)
                            constatacion=ConstatacionFisica(usuariobienes=activo.responsable,
                                                            # numero=secuencia.numeroconstatacion,
                                                            # normativaconstatacion=secuencia.normativaconstatacion,
                                                            fechainicio=hoy,
                                                            periodo_id=id_p,
                                                            ubicacionbienes=activo.ubicacion)
                            constatacion.save(request)
                            log(u'Agrego constatación: %s' % (constatacion), request, "add")
                        detalle_c=DetalleConstatacionFisica(codigoconstatacion=constatacion,
                                                            activo=activo,
                                                            responsable=persona,
                                                            estadooriginal=activo.estado,
                                                            )
                        detalle_c.save(request)
                        log(u'Agrego detalle de constatación: %s' % (detalle_c), request, "add")
                    encontrado = eval(request.POST['val'].capitalize())
                    detalle_c.encontrado = encontrado
                    detalle_c.enuso = encontrado
                    detalle_c.perteneceusuario = True
                    if not encontrado:
                        detalle_c.requieretraspaso = False
                        detalle_c.requieredarbaja = False
                        detalle_c.estadoactual = None
                        detalle_c.condicionestado = None
                        detalle_c.ubicacionbienes = None
                        detalle_c.usuariobienes = None
                    else:
                        detalle_c.estadoactual = activo.estado
                        detalle_c.condicionestado = activo.condicionestado
                        detalle_c.ubicacionbienes = activo.ubicacion
                        detalle_c.usuariobienes = activo.responsable
                    detalle_c.save(request)

                    if detalle_c.enuso != activo.enuso:
                        historial_estado = HistorialEstadoActivo(activo=activo, persona=persona,
                                                                 estado=activo.estado, enuso=detalle_c.enuso,
                                                                 observacion='Registrado desde el botón el check encontrado de constatación de activos fijos',
                                                                 tiporegistro=5, condicionestado=activo.condicionestado)
                        historial_estado.save(request)
                        activo.enuso=detalle_c.enuso
                        activo.save(request)
                    context={'id':encrypt(detalle_c.id),
                             'id_a':encrypt(detalle_c.activo.id),
                             'id_p':encrypt(detalle_c.codigoconstatacion.periodo.id),
                             'enuso':detalle_c.enuso,
                             'requieretraspaso':detalle_c.requieretraspaso,
                             'requieredarbaja':detalle_c.requieredarbaja,
                             'perteneceusuario':detalle_c.perteneceusuario,
                             }
                    log(u'Edito estado encontrado y en uso de detalle constatación: %s (%s)' % (detalle_c, detalle_c.encontrado), request,"edit")
                    return JsonResponse({"result": True, 'mensaje': 'Cambios guardados','context':context})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': False, "mensaje": 'Error: {}'.format(ex)}, safe=False)

        elif action == 'constatar':
            try:
                form = ConstatacionFisicaForm(request.POST)
                encontrado=request.POST.get('encontrado','')
                if not encontrado:
                    form.fields['estadoactual'].required=False
                    form.fields['condicionestado'].required=False
                if not form.is_valid():
                    transaction.set_rollback(True)
                    form_error = [{k: v[0]} for k, v in form.errors.items()]
                    return JsonResponse({'result': True, "form": form_error, "mensaje": "Error en el formulario"})
                id_a, id_p, id_d =encrypt_id(request.POST['ida']), encrypt_id(request.POST['idp']), encrypt_id(request.POST['id'])
                estado, condicionestado, enuso = form.cleaned_data['estadoactual'], \
                                                 form.cleaned_data['condicionestado'],\
                                                 form.cleaned_data['enuso']
                detalle_c = DetalleConstatacionFisica.objects.filter(status=True, activo_id=id_a, codigoconstatacion__periodo_id=id_p, codigoconstatacion__status=True).first()
                activo = ActivoFijo.objects.get(id=id_a)
                condicionestado = int(condicionestado) if condicionestado else activo.condicionestado
                estado = estado if estado else activo.estado

                if form.cleaned_data['requieretraspaso']:
                    usuariobienes = form.cleaned_data['usuariobienes']
                elif form.cleaned_data['encontrado']:
                    usuariobienes = activo.responsable
                else:
                    usuariobienes = None
                if not detalle_c:
                    constatacion = ConstatacionFisica.objects.filter(status=True, usuariobienes=activo.responsable, periodo_id=id_p).first()
                    if not constatacion:
                        # secuencia = secuencia_activos(request)
                        # secuencia.numeroconstatacion += 1
                        # secuencia.save(request)
                        constatacion = ConstatacionFisica(usuariobienes=activo.responsable,
                                                          # numero=secuencia.numeroconstatacion,
                                                          # normativaconstatacion=secuencia.normativaconstatacion,
                                                          fechainicio=hoy,
                                                          periodo_id=id_p,
                                                          ubicacionbienes=activo.ubicacion)
                        constatacion.save(request)
                        log(u'Agrego constatación: %s' % (constatacion), request, "add")
                    detalle_c = DetalleConstatacionFisica(codigoconstatacion=constatacion,
                                                          activo=activo,
                                                          responsable=persona,
                                                          usuariobienes=usuariobienes,
                                                          ubicacionbienes=form.cleaned_data['ubicacionbienes'],
                                                          estadooriginal=activo.estado,
                                                          estadoactual=form.cleaned_data['estadoactual'],
                                                          condicionestado=form.cleaned_data['condicionestado'] if form.cleaned_data['condicionestado'] else None,
                                                          enuso=form.cleaned_data['enuso'],
                                                          encontrado=form.cleaned_data['encontrado'],
                                                          requieretraspaso=form.cleaned_data['requieretraspaso'],
                                                          requieredarbaja=form.cleaned_data['requieredarbaja'],
                                                          perteneceusuario=True,
                                                          observacion=form.cleaned_data['observacion'])
                    detalle_c.save(request)
                    log(u'Agrego detalle de constatación: %s' % (detalle_c), request, "add")
                else:
                    if not detalle_c.responsable.id == persona.id:
                        raise NameError(f'Activo que intenta constatar ya fue constatado por {detalle_c.responsable.nombre_completo_minus()}')
                    detalle_c.usuariobienes = usuariobienes
                    detalle_c.ubicacionbienes = form.cleaned_data['ubicacionbienes']
                    detalle_c.estadooriginal = activo.estado
                    detalle_c.estadoactual=form.cleaned_data['estadoactual']
                    detalle_c.condicionestado=form.cleaned_data['condicionestado'] if form.cleaned_data['condicionestado'] else None
                    detalle_c.enuso = form.cleaned_data['enuso']
                    detalle_c.encontrado = form.cleaned_data['encontrado']
                    detalle_c.requieretraspaso = form.cleaned_data['requieretraspaso']
                    detalle_c.requieredarbaja = form.cleaned_data['requieredarbaja']
                    # detalle_c.perteneceusuario = form.cleaned_data['perteneceusuario']
                    detalle_c.observacion = form.cleaned_data['observacion']
                    detalle_c.save(request)
                    log(f'Edito detalle de constatación de activo fijo: {detalle_c}', request, 'edit')
                if encontrado and (estado != activo.estado or condicionestado != activo.condicionestado or form.cleaned_data['enuso'] != activo.enuso):
                    historial_estado = HistorialEstadoActivo(activo=activo, persona=persona,
                                                             estado=estado, enuso=form.cleaned_data['enuso'],
                                                             observacion=form.cleaned_data['observacion'],
                                                             tiporegistro=5, condicionestado=int(condicionestado))
                    historial_estado.save(request)
                    activo.estado = form.cleaned_data['estadoactual']
                    activo.condicionestado = condicionestado
                    activo.estado = estado
                    activo.enuso = form.cleaned_data['enuso']
                    activo.save(request)
                    log(f'Edito estado y proceso de bja de activo fijo: {activo}', request, 'edit')
                context = {'id': encrypt(detalle_c.id),
                           'id_a': encrypt(detalle_c.activo.id),
                           'id_p': encrypt(detalle_c.codigoconstatacion.periodo.id),
                           'enuso': detalle_c.enuso,
                           'encontrado': detalle_c.encontrado,
                           'requieretraspaso': detalle_c.requieretraspaso,
                           'requieredarbaja': detalle_c.requieredarbaja,
                           'perteneceusuario': detalle_c.perteneceusuario,
                           'estado': activo.estado.nombre,
                           'modal': True,
                           }
                return JsonResponse({'result': True, 'mensaje': 'Guardado con éxito', 'data_return':True, 'data':context})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': str(ex)})

        elif action == 'delconstatacion':
            try:
                detalle_c = DetalleConstatacionFisica.objects.get(id=encrypt_id(request.POST['id']))
                detalle_c.status = False
                detalle_c.save(request)
                log(f'Elimino constatacion de activos fijos: {detalle_c}', request, 'del')
                res_json = {"error": False}
            except Exception as ex:
                res_json = {'error': True, "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)

        elif action == 'activarperiodo':
            with transaction.atomic():
                try:
                    periodo = PeriodoConstatacionAF.objects.get(pk=request.POST['id'])
                    periodo.activo = eval(request.POST['val'].capitalize())
                    periodo.save(request)
                    if periodo.activo:
                        periodos = PeriodoConstatacionAF.objects.filter(status=True, activo=True).exclude(id=periodo.id)
                        periodos.update(activo=False)
                    log(u'Configuración requisitos activo : %s (%s)' % (periodo, periodo.activo), request, "edit")
                    return JsonResponse({"result": True, 'mensaje': 'Cambios guardados'})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": False})

        # ARREGLO INFORMES DE BAJA
        elif action == 'arregloinformesbajaold':
            try:
                director = get_directorresponsablebaja('AF')
                documentos = DocumentoFirmaInformeBaja.objects.filter(status=True, informe__status=True,
                                                                      informe__tipoinforme=2, estadofirma=2,
                                                                      responsablefirma__isnull=False).order_by('-id')

                if not documentos:
                    return JsonResponse({"result": 'bad', "mensaje": "No hay documentos para procesar"})
                for index, documento in enumerate(documentos):
                    informe = documento.informe

                    activo = informe.activofijo
                    activo.archivobaja = None
                    activo.estado_id = 3
                    activo.condicionestado = 1
                    activo.enuso = False
                    activo.save(request)

                    historial_activo = HistorialEstadoActivo(activo=activo, estado=activo.estado,
                                                             condicionestado=activo.condicionestado,
                                                             persona_id=1,
                                                             observacion='Actualización de informe baja masivo',
                                                             enuso=activo.enuso,
                                                             tiporegistro=1)
                    historial_activo.save(request)

                    informe.estadoactivo = activo.estado
                    informe.enuso = activo.enuso
                    informe.estado = activo.condicionestado
                    informe.fecha_creacion = datetime.now()
                    informe.save(request)

                    file_obj, response = generar_informebaja_pdf_v2(informe, director)
                    documento.archivo = file_obj
                    documento.director = director
                    documento.firmadirector = False
                    documento.estadofirma = 1
                    documento.save(request)

                    historial = HistorialDocumentoInformeBaja(
                        documentoinforme=documento,
                        persona_id=1,
                        archivo=documento.archivo,
                        estadofirma=4
                    )
                    historial.save(request)
                    print(f'Procesado {index + 1} de {documentos.count()}. Id de documento: {documento.id}. Success')

                return JsonResponse({"result": 'ok', "mensaje": "Proceso finalizado"})
            except Exception as ex:
                print(f'Error: {ex}')
                transaction.set_rollback(True)
                return JsonResponse({"result": 'bad', "mensaje": str(ex)})

        # INFORME DE BAJA
        elif action == 'addinformebaja':
            try:
                viewinformesbaja = request.POST.get('val_extra', '')
                form = InformeBajaFormAF(request.POST)
                activo = ActivoFijo.objects.get(id=encrypt_id(request.POST['id']))
                if not form.is_valid():
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()],
                                         "mensaje": "Error en el formulario"})
                informebaja = activo.existeinformebaja()
                estadoactivo = form.cleaned_data['estadoactivo']
                enuso = form.cleaned_data['enuso']
                condicionestado = form.cleaned_data['estado']
                if not informebaja:
                    informebaja = InformeActivoBaja(activofijo=activo,
                                                    solicita=form.cleaned_data['solicita'],
                                                    responsable=form.cleaned_data['responsable'],
                                                    departamento=form.cleaned_data['departamento'],
                                                    gestion=form.cleaned_data['gestion'],
                                                    conclusion=form.cleaned_data['conclusion'],
                                                    estado=condicionestado,
                                                    estadouso=form.cleaned_data['estadouso'],
                                                    bloque=form.cleaned_data['bloque'],
                                                    detallerevision=form.cleaned_data['detallerevision'],
                                                    estadoactivo=estadoactivo,
                                                    enuso=enuso,
                                                    tipoinforme=2)
                    informebaja.save(request)
                    log(u'Adiciono informe activo de baja: %s' % informebaja, request, "add")
                else:
                    informebaja.solicita = form.cleaned_data['solicita']
                    informebaja.responsable = form.cleaned_data['responsable']
                    informebaja.conclusion = form.cleaned_data['conclusion']
                    informebaja.departamento = form.cleaned_data['departamento']
                    informebaja.gestion = form.cleaned_data['gestion']
                    informebaja.estado = condicionestado
                    informebaja.estadouso = form.cleaned_data['estadouso']
                    informebaja.bloque = form.cleaned_data['bloque']
                    informebaja.detallerevision = form.cleaned_data['detallerevision']
                    informebaja.estadoactivo = estadoactivo
                    informebaja.enuso = enuso
                    informebaja.save(request)
                    log(u'Edito informe activo de baja: %s' % informebaja, request, "edit")
                if 'actividades' in request.POST:
                    listadetalle = request.POST.getlist('actividades')
                    if listadetalle:
                        informebaja.actividades_informe_baja().update(status=False)
                        for lisdet in listadetalle:
                            ingresodetalle = DetalleInformeActivoBaja(informebaja=informebaja, detalle=lisdet)
                            ingresodetalle.save(request)

                if activo.estado != estadoactivo or activo.condicionestado != int(condicionestado) or activo.enuso != enuso:
                    activo.estado = estadoactivo
                    activo.enuso = enuso
                    activo.condicionestado = condicionestado
                    activo.save(request)
                    historial = HistorialEstadoActivo(activo=activo, estado=activo.estado,
                                                      condicionestado=activo.condicionestado,
                                                      observacion=form.cleaned_data['detallerevision'],
                                                      persona=persona,
                                                      enuso=activo.enuso,
                                                      tiporegistro=1)
                    historial.save(request)

                documento = informebaja.documento_informe_baja()
                director = get_directorresponsablebaja('AF')

                file_obj, response = generar_informebaja_pdf_v2(informebaja, director)

                if not documento:
                    documento = DocumentoFirmaInformeBaja(informe=informebaja,
                                                          director=director,
                                                          estadofirma=1,
                                                          archivo=file_obj)
                else:
                    documento.director = director
                    documento.informe = informebaja
                    documento.estadofirma = 1
                    documento.archivo = file_obj
                documento.save(request)

                if 'firmarinforme' in request.POST:
                    file_obj = firmar_doc_informebaja(request, documento)
                    documento.archivo = file_obj
                    documento.estadofirma = 2
                    documento.save(request)

                historial = HistorialDocumentoInformeBaja(
                    documentoinforme=documento,
                    persona=persona,
                    archivo=documento.archivo,
                    estadofirma=documento.estadofirma
                )
                historial.save(request)
                log(u'Guardo historial de archivo firmado: {}'.format(historial), request, "add")
                codigo = activo.codigogobierno if activo.codigogobierno else activo.codigointerno
                if viewinformesbaja:
                    return JsonResponse({"result": False, 'to': f'{request.path}?action=informesbaja&s={codigo}'}, safe=True)
                else:
                    return JsonResponse({"result": False, 'to': f'{request.path}?s={codigo}'}, safe=True)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": str(ex)})

        elif action == 'adddirectorresponsablebaja':
            try:
                form = ResponsableInformeBajaForm(request.POST)
                id = encrypt_id(request.POST['id'])
                director = DirectorResponsableBaja.objects.filter(pk=int(id)).first()
                if not form.is_valid():
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()],
                                         "mensaje": "Error en el formulario"})
                if not director:
                    director = DirectorResponsableBaja(responsable=form.cleaned_data['responsable'],
                                                          cargo=form.cleaned_data['cargo'],
                                                       fechainicio=form.cleaned_data['fechainicio'],
                                                       fechafin=form.cleaned_data['fechafin'],
                                                       departamento=form.cleaned_data['departamento'],
                                                       actual=form.cleaned_data['actual'])

                else:
                    director.responsable = form.cleaned_data['responsable']
                    director.cargo = form.cleaned_data['cargo']
                    director.fechainicio = form.cleaned_data['fechainicio']
                    director.fechafin = form.cleaned_data['fechafin']
                    director.departamento=form.cleaned_data['departamento']
                    director.actual = form.cleaned_data['actual']
                director.save(request)
                if director.actual:
                    DirectorResponsableBaja.objects.filter(status=True, departamento=director.departamento).exclude(id=director.id).update(actual=False)
                log(u'Adiciono director responsable de baja: %s' % director, request, "add")
                return JsonResponse({"result": False, 'mensaje': 'Guardado con éxito'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": str(ex)})


        elif action == 'verificaciontecnica':
            try:
                form = VerificacionTecnicaForm(request.POST)
                activo = ActivoFijo.objects.get(id=encrypt_id(request.POST['id']))
                if not form.is_valid():
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()],
                                         "mensaje": "Error en el formulario"})

                activo.estado = form.cleaned_data['estadoactivo']
                activo.condicionestado = form.cleaned_data['condicionestado']
                activo.enuso = form.cleaned_data['enuso']
                activo.save(request)

                historial = HistorialEstadoActivo(activo=activo, estado=activo.estado,
                                                  condicionestado=activo.condicionestado,
                                                  observacion=form.cleaned_data['observacion'],
                                                  persona=persona,
                                                  enuso=activo.enuso,
                                                  tiporegistro=2)
                historial.save(request)

                # nombre_archivo = generar_nombre(f'informeverificacion_{historial.id}_', 'generado') + '.pdf'
                # template_path = "af_activofijo/informeverificaciontecnica.html"
                # context = {
                #     'pagesize': 'A4',
                #     'historial': historial,
                #     'hoy': datetime.now().date()
                # }
                # pdf_file, response = conviert_html_to_pdf_save_file_model(template_path, context, nombre_archivo)
                # historial.informe = pdf_file
                # historial.save(request)

                codigo = activo.codigogobierno if activo.codigogobierno else activo.codigointerno
                return JsonResponse({"result": False, 'to': f'{request.path}?s={codigo}'}, safe=True)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": str(ex)})


        elif action == 'delinformebaja':
            with transaction.atomic():
                try:
                    instancia = InformeActivoBaja.objects.get(pk=int(encrypt(request.POST['id'])))
                    instancia.status = False
                    instancia.save(request)
                    instancia.documentofirmainformebaja_set.filter(status=True).update(status=False)
                    log(u'Elimino informe de activo baja y sus dependencias: %s' % instancia, request, "del")
                    res_json = {"error": False}
                except Exception as ex:
                    res_json = {'error': True, "message": "Error: {}".format(ex)}
                return JsonResponse(res_json, safe=False)

        elif action == 'firmarinformebaja':
            try:
                # FIRMA ELECTRÓNICA
                certificado = request.FILES["firma"]
                contrasenaCertificado = request.POST['palabraclave']
                razon = request.POST['razon'] if 'razon' in request.POST else ''
                extension_certificado = os.path.splitext(certificado.name)[1][1:]
                bytes_certificado = certificado.read()
                firmadirector = request.POST.get('val_extra', '')

                # PARÁMETROS
                id = encrypt_id(request.POST['id'])
                director = get_directorresponsablebaja('AF')

                if not firmadirector:
                    informe = InformeActivoBaja.objects.get(id=id)
                    documento = informe.documento_informe_baja()
                    # archivo_, archivo_url = generar_informebaja(request, informe)
                    ruta_hoja_membretada = get_ruta_hoja_membretada()
                    perso = informe.responsable
                    detalle = informe.detalleinformeactivobaja_set.filter(status=True)
                    nombre_archivo = generar_nombre(f'Informe_Baja{informe.activofijo.id}_', 'generado') + '.pdf'
                    template_path = 'af_activofijo/informebajapdfmantenimiento.html'
                    context = {'pagesize': 'A4', 'ruta_hoja_membretada': ruta_hoja_membretada,
                               'informe': informe.activofijo, 'perso': perso, 'director': director,
                               'bajaactivo': informe, 'detalle': detalle, 'hoy': datetime.now().date()
                               }
                    file_obj, response = conviert_html_to_pdf_save_file_model(template_path, context, nombre_archivo)

                    if not documento:
                        documento = DocumentoFirmaInformeBaja(informe=informe,
                                                              director=director,
                                                              estadofirma=2,
                                                              archivo=file_obj)
                    else:
                        documento.archivo = file_obj
                        documento.estadofirma = 2
                        documento.director = director
                    documento.save(request)
                    archivo_ = documento.archivo
                    archivo_url = archivo_.url
                else:
                    documento = DocumentoFirmaInformeBaja.objects.get(id=id)
                    informe = documento.informe
                    archivo_ = documento.ultimo_dcoumento().archivo
                    archivo_url = archivo_.url
                # director = director_af(informe)

                # VALIDACIONES
                if firmadirector and not director.responsable.id == persona.id:
                    raise NameError("Para poder firmar tiene que ser el director de departamento")
                if not firmadirector and not informe.responsable.id == persona.id:
                    raise NameError("Para poder firmar tiene que ser el responsable del informe de baja")

                cargo = persona.mi_cargo_administrativo() if persona.mi_cargo_administrativo() else ''
                palabras = f'{persona} {cargo}'
                x, y, numPage = obtener_posicion_x_y_saltolinea(archivo_url, palabras, False, True)
                if not x or not y:
                    raise NameError('No se encontró el responsable en el documento.')
                datau = JavaFirmaEc(archivo_a_firmar=archivo_, archivo_certificado=bytes_certificado,
                                    extension_certificado=extension_certificado,
                                    password_certificado=contrasenaCertificado,
                                    page=int(numPage), reason=razon, lx=x, ly=y).sign_and_get_content_bytes()
                archivo_ = io.BytesIO()
                archivo_.write(datau)
                archivo_.seek(0)

                _name = f"InformeBaja_{informe.id}"
                file_obj = DjangoFile(archivo_, name=f"{_name}.pdf")
                documento.firmadirector = True if firmadirector else False
                documento.archivo = file_obj
                documento.save(request)
                log(u'Edito archivo firmado: {}'.format(documento), request, "edit")
                if firmadirector:
                    documento.estadofirma = 3
                    documento.save(request)
                    activofijo = informe.activofijo
                    activofijo.archivobaja = file_obj
                    activofijo.save(request)
                    experto = Persona.objects.get(id=1204)
                    titulo = f'Informe de baja registrado {activofijo.id}'
                    cuerpo = f'Se registro un nuevo informe de baja del activo {activofijo}'
                    notify(titulo, cuerpo + ' a su nombre.', activofijo.responsable, None,
                           f'/mis_activos?s={activofijo.codigogobierno}', activofijo.pk, 2, 'sga-sagest',
                           ActivoFijo, request)
                    notify(titulo, cuerpo, experto, None,
                           f'/af_activofijo?action=informesbaja&s={activofijo.codigogobierno}', activofijo.pk, 2,
                           'sga-sagest', ActivoFijo, request)
                    log(u'Edito archivo firmado: {}'.format(activofijo), request, "add")
                historial = HistorialDocumentoInformeBaja(
                    documentoinforme=documento,
                    persona=persona,
                    archivo=file_obj,
                    estadofirma=2
                )
                historial.save(request)

                log(u'Guardo historial de archivo firmado: {}'.format(historial), request, "add")
                return JsonResponse({"result": False, "mensaje": "Guardado con exito", }, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": u"Error al guardar. %s" % ex.__str__()}, safe=False)

        elif action == 'firmarinformebajamasivo':
            try:
                # FIRMA ELECTRÓNICA
                certificado = request.FILES["firma"]
                contrasenaCertificado = request.POST['palabraclave']
                razon = request.POST['razon'] if 'razon' in request.POST else ''
                extension_certificado = os.path.splitext(certificado.name)[1][1:]
                bytes_certificado = certificado.read()
                firmadirector = request.POST.get('cond_extra', '')
                director = get_directorresponsablebaja('AF')
                limit = int(request.POST['val_extra'])
                if firmadirector:
                    director = get_directorresponsablebaja('AF')
                    documentos = DocumentoFirmaInformeBaja.objects.filter(status=True, firmadirector=False,director=director, estadofirma=2, informe__tipoinforme=2)[:limit]
                else:
                    documentos = DocumentoFirmaInformeBaja.objects.filter(status=True, firmadirector=False, informe__responsable=persona, estadofirma=1, informe__tipoinforme=2)[:limit]
                for documento in documentos:
                    try:
                        informe = documento.informe
                        archivo_ = documento.ultimo_dcoumento().archivo
                        archivo_url = archivo_.url
                        # director = director_af(informe)
                        cargo = persona.mi_cargo_administrativo() if persona.mi_cargo_administrativo() else ''
                        palabras = f'{persona} {cargo}'
                        x, y, numPage = obtener_posicion_x_y_saltolinea(archivo_url, palabras, False, True)
                        if x and y:
                            datau = JavaFirmaEc(archivo_a_firmar=archivo_, archivo_certificado=bytes_certificado,
                                                extension_certificado=extension_certificado,
                                                password_certificado=contrasenaCertificado,
                                                page=int(numPage), reason=razon, lx=x, ly=y).sign_and_get_content_bytes()
                            archivo_ = io.BytesIO()
                            archivo_.write(datau)
                            archivo_.seek(0)

                            _name = f"InformeBaja_{informe.id}"
                            file_obj = DjangoFile(archivo_, name=f"{_name}.pdf")
                            documento.firmadirector = True if firmadirector else False
                            documento.archivo = file_obj
                            documento.estadofirma = 3 if firmadirector else 2
                            if not firmadirector:
                                documento.director = director
                            documento.save(request)
                            log(u'Edito archivo firmado: {}'.format(documento), request, "edit")
                            if firmadirector:
                                activofijo = informe.activofijo
                                activofijo.archivobaja = file_obj
                                activofijo.save(request)
                                experto = Persona.objects.get(id=1204)
                                titulo = f'Informe de baja registrado {activofijo.id}'
                                cuerpo = f'Se registro un nuevo informe de baja del activo {activofijo}'
                                notify(titulo, cuerpo + ' a su nombre.', activofijo.responsable, None, f'/mis_activos?s={activofijo.codigogobierno}', activofijo.pk, 2, 'sga-sagest', ActivoFijo, request)
                                notify(titulo, cuerpo, experto, None, f'/af_activofijo?action=informesbaja&s={activofijo.codigogobierno}', activofijo.pk, 2, 'sga-sagest', ActivoFijo, request)
                                log(u'Edito archivo firmado: {}'.format(activofijo), request, "add")
                            historial = HistorialDocumentoInformeBaja(
                                documentoinforme=documento,
                                persona=persona,
                                archivo=file_obj,
                                estadofirma=2
                            )
                            historial.save(request)
                            log(u'Guardo historial de archivo firmado: {}'.format(historial), request, "add")
                    except Exception as ex:
                        if f'{ex}' == 'Certificado no es válido' or f'{ex}' == 'Invalid password or PKCS12 data':
                            raise NameError(f'{ex}')
                return JsonResponse({"result": False, "mensaje": "Guardado con exito", }, safe=False)
            except Exception as ex:
                print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
                textoerror = '{} Linea:{}'.format(ex, sys.exc_info()[-1].tb_lineno)
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error: %s" % ex})

        elif action == 'addarchivobaja':
            try:
                form = ArchivoActivoBajaForm(request.POST, request.FILES)
                if form.is_valid():
                    activo = ActivoFijo.objects.get(pk=encrypt(request.POST['id']))
                    newfile = request.FILES['archivobaja']
                    newfile._name = generar_nombre(f"InformeBaja_{activo.id}", newfile._name)
                    informe = activo.existeinformebaja()
                    documento = informe.documento_informe_baja()
                    if not documento:
                        # director = director_af(informe)
                        # if not director:
                        #     raise NameError('Sin director configurado.')
                        director = get_directorresponsablebaja('AF')
                        documento = DocumentoFirmaInformeBaja(informe=informe,
                                                              director=director,
                                                              estadofirma=2,
                                                              archivo=newfile)
                        documento.save(request)
                        log(u'Guardo archivo firmado: {}'.format(documento), request, "add")
                    else:
                        documento.firmadirector = False
                        documento.archivo = newfile
                        documento.estadofirma = 2
                        documento.save(request)
                        log(u'Edito archivo firmado: {}'.format(documento), request, "edit")
                        # if firmadirector:
                        #     activofijo = informe.activofijo
                        #     activofijo.archivobaja = file_obj
                        #     activofijo.save(request)
                        #     log(u'Edito archivo firmado: {}'.format(activofijo), request, "add")
                    historial = HistorialDocumentoInformeBaja(
                        documentoinforme=documento,
                        persona=persona,
                        estadofirma=3,
                        archivo=newfile
                    )
                    historial.save(request)
                    return JsonResponse({"result": False}, safe=False)
                else:
                    raise NameError('Debe subir archivo pdf con los MG especificado')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": u"Error al guardar los datos. {}".format(ex)})

        elif action == 'addinformebajafirmado':
            try:
                form = ArchivoActivoBajaForm(request.POST, request.FILES)
                if form.is_valid():
                    activo = ActivoFijo.objects.get(pk=encrypt(request.POST['id']))
                    newfile = request.FILES['archivobaja']
                    newfile._name = generar_nombre(f"InformeBaja_{activo.id}", newfile._name)
                    activo.archivobaja = newfile
                    activo.save(request)
                    return JsonResponse({"result": False}, safe=False)
                else:
                    raise NameError('Debe subir archivo pdf con los MG especificado')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": u"Error al guardar los datos. "})

        elif action == 'delacta':
            with transaction.atomic():
                try:
                    instancia = DocumentoFirmaInformeBaja.objects.get(pk=int(encrypt(request.POST['id'])))
                    instancia.status = False
                    instancia.save(request)

                    InformeActivoBaja.objects.filter(id=instancia.informe.id).update(status=False)

                    activo = instancia.informe.activofijo
                    if activo:
                        activo.archivobaja = None
                        activo.save(request)
                    log(u'Elimino acta de constatación: %s' % instancia, request, "del")
                    res_json = {"error": False}
                except Exception as ex:
                    res_json = {'error': True, "message": "Error: {}".format(ex)}
                return JsonResponse(res_json, safe=False)

        elif action == 'constatacionmasiva':
            try:
                archivo_ = request.FILES['archivo']
                idperiodo = encrypt_id(request.POST['id_obj'])
                nombres_hojas = pd.ExcelFile(archivo_).sheet_names
                cabecera = ['codigogobierno', 'codigointerno',
                            'estado','condicionestado', 'cedulaconstatador',
                            'responsableactivo', 'cedularesponsable',
                            'codigoubicacion', 'ubicacion', 'observacion']
                total_registros, cont= 0, 0
                for name in nombres_hojas:
                    df = pd.read_excel(archivo_, sheet_name=name)
                    df.columns = formatear_cabecera_pd(df)
                    for c in cabecera:
                        if not c in df.columns:
                            raise NameError(f'Formato de archivo erróneo: La columna {c} no se encuentra en el documento.')

                    for index, row in df.iterrows():
                        codigogobierno = str(row['codigogobierno']).strip().split('.')[0]
                        codigointerno = str(row['codigointerno']).strip().split('.')[0]
                        codigoubicacion = str(row['codigoubicacion']).strip()
                        condicionestado = str(row['condicionestado']).strip()
                        cedulaconstatador = str(row['cedulaconstatador']).strip().split('.')[0]
                        cedularesponsable = str(row['cedularesponsable']).strip().split('.')[0]
                        responsableactivo = str(row['responsableactivo']).strip().lower()
                        observacion = str(row['observacion']).strip()
                        ubicacion = str(row['ubicacion']).strip().lower()
                        estado = str(row['estado']).strip().upper()
                        cedulaconstatador = f'0{cedulaconstatador}' if len(cedulaconstatador) == 9 else cedulaconstatador
                        cedularesponsable = f'0{cedularesponsable}' if len(cedularesponsable) == 9 else cedularesponsable
                        activo = ActivoFijo.objects.filter(Q(codigogobierno__iexact=codigogobierno) |
                                                           Q(codigointerno__iexact=codigogobierno)).first()

                        if codigointerno and not activo:
                            activo = ActivoFijo.objects.filter(Q(codigointerno__iexact=codigointerno) |
                                                               Q(codigogobierno__iexact=codigointerno)).first()

                        if activo:
                            if codigoubicacion:
                                ubicacion = Ubicacion.objects.filter(codigo=codigoubicacion).first()
                            else:
                                ubicacion = Ubicacion.objects.filter(nombre__unaccent__iexact=ubicacion).first()
                            constatador = Persona.objects.filter(cedula=cedulaconstatador, status=True).first()
                            usuariobienes = Persona.objects.filter(cedula=cedularesponsable, status=True).first()
                            usuariobienes = activo.responsable if not usuariobienes else usuariobienes
                            estadoactual = EstadoProducto.objects.filter(nombre__unaccent__iexact=estado).first()
                            estadoactual = activo.estado if not estadoactual else estadoactual
                            condicionestado = condicionestado if condicionestado and condicionestado != 'nan' else None
                            if estadoactual:
                                id_estado = estadoactual.id
                                if id_estado == 1:
                                    condicionestado = 4
                                elif id_estado == 3:
                                    condicionestado = 1
                                elif id_estado == 2 and condicionestado:
                                    condicionestado = 2 if unidecode(condicionestado).upper() == 'OBSOLETO' else 3
                            detalle_c = DetalleConstatacionFisica.objects.filter(status=True, activo=activo, codigoconstatacion__periodo_id=idperiodo).first()
                            if not detalle_c:
                                constatacion = ConstatacionFisica.objects.filter(status=True, usuariobienes=activo.responsable, periodo_id=idperiodo).first()
                                if not constatacion:
                                    secuencia = secuencia_activos(request)
                                    secuencia.numeroconstatacion += 1
                                    secuencia.save(request)
                                    constatacion = ConstatacionFisica(usuariobienes=activo.responsable,
                                                                      numero=secuencia.numeroconstatacion,
                                                                      normativaconstatacion=secuencia.normativaconstatacion,
                                                                      fechainicio=hoy,
                                                                      periodo_id=idperiodo,
                                                                      ubicacionbienes=activo.ubicacion)
                                    constatacion.save(request)
                                    log(u'Agrego constatación: %s' % (constatacion), request, "add")
                                detalle_c = DetalleConstatacionFisica(codigoconstatacion=constatacion,
                                                                      activo=activo,
                                                                      responsable=constatador,
                                                                      usuariobienes=usuariobienes,
                                                                      ubicacionbienes=ubicacion,
                                                                      estadooriginal=activo.estado,
                                                                      estadoactual=estadoactual,
                                                                      observacion=observacion,
                                                                      condicionestado=condicionestado,
                                                                      enuso=True,
                                                                      encontrado=True)
                                detalle_c.save(request)
                                log(u'Agrego detalle de constatación: %s' % (detalle_c), request, "add")
                                if activo.estado != estadoactual:
                                    activo.estado = estadoactual
                                    activo.save(request)
                                    log(f'Edito estado y proceso de bja de activo fijo: {activo}', request, 'edit')
                                if condicionestado and condicionestado != 'nan' and activo.condicionestado != condicionestado:
                                    activo.condicionestado = condicionestado
                                    activo.save(request)
                            # else:
                            #     cont += 1
                        else:
                           cont += 1
                    total_registros += df.shape[0]
                total_registros = total_registros-cont
                messages.success(request, f'Se constato {total_registros} registros, Registros no constatados: {cont}')
                return JsonResponse({"result":False, 'mensaje':'Importado correctamente'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": f"Error: {ex}"})

        elif action == 'reportehistoricoactivosconstatacion':
            try:
                if 'fechadesde' not in request.POST or 'fechahasta' not in request.POST or 'codigo' not in request.POST:
                    return JsonResponse({"result": "bad", "titulo": "Error", "mensaje": u"Error al procesar la solicitud", "showSwal": "True", "swalType": "error"})

                fechadesde = datetime.strptime(request.POST['fechadesde'], '%Y-%m-%d').date()
                fechahasta = datetime.strptime(request.POST['fechahasta'], '%Y-%m-%d').date()
                codigo = request.POST['codigo']

                if not codigo:
                    if not ConstatacionFisica.objects.values("id").filter(status=True, estado=2, fechainicio__range=[fechadesde, fechahasta], fechafin__range=[fechadesde, fechahasta]).exists():
                        return JsonResponse({"result": "bad", "titulo": "Atención!!!", "mensaje": "No existen Registros de constataciones físicas con estado <b>FINALIZADO</b> en ese rango de fechas", "showSwal": "True", "swalType": "warning"})
                else:
                    if not ConstatacionFisica.objects.values("id").filter(Q(detalleconstatacionfisica__activo__codigogobierno__icontains=codigo) | Q(detalleconstatacionfisica__activo__codigointerno__icontains=codigo),status=True, estado=2, fechainicio__range=[fechadesde, fechahasta], fechafin__range=[fechadesde, fechahasta]).exists():
                        return JsonResponse({"result": "bad", "titulo": "Atención!!!", "mensaje": "No existen Registros de constataciones físicas con estado <b>FINALIZADO</b> en ese rango de fechas para el activo", "showSwal": "True", "swalType": "warning"})

                # Guardar la notificación
                notificacion = Notificacion(
                    cuerpo='Generación de reporte de excel en progreso',
                    titulo='Reporte Excel Histórico Estado de Activos por Constataciones Física',
                    destinatario=persona,
                    url='',
                    prioridad=1,
                    app_label='SGA',
                    fecha_hora_visible=datetime.now() + timedelta(days=1),
                    tipo=2,
                    en_proceso=True
                )
                notificacion.save(request)

                reporte_historico_estado_activos_por_constatacion_background(request=request, data=data, idnotificacion=notificacion.id, fechadesde=fechadesde, fechahasta=fechahasta, codigo=codigo).start()

                return JsonResponse({"result": "ok",
                                     "mensaje": u"El reporte se está generando. Verifique su apartado de notificaciones después de unos minutos.",
                                     "btn_notificaciones": traerNotificaciones(request, data, persona)})
            except Exception as ex:
                msg = ex.__str__()
                return JsonResponse({"result": "bad", "titulo": "Error", "mensaje": u"Error al generar el reporte. [%s]" % msg, "showSwal": "True", "swalType": "error"})

        elif action == 'firmaractaconstatacion':
            try:
                id = encrypt_id(request.POST['id'])
                constatacion = ConstatacionFisica.objects.get(pk=id)
                if not constatacion.acta_persona_firmada(persona):
                    documento_a_firmar = constatacion.get_documento()
                    certificado = request.FILES["firma"]
                    contrasenaCertificado = request.POST['palabraclave']
                    razon = request.POST['razon'] if 'razon' in request.POST else ''
                    jsonFirmas = json.loads(request.POST['txtFirmas'])
                    _name, extension_documento_a_firmar = f'acta_constatacion_{constatacion.usuariobienes.usuario.username}', '.pdf'
                    extension_certificado = os.path.splitext(certificado.name)[1][1:]
                    bytes_certificado = certificado.read()
                    if not jsonFirmas:
                        raise NameError("Debe seleccionar ubicación de la firma")
                    for membrete in jsonFirmas:
                        datau = JavaFirmaEc(
                            archivo_a_firmar=documento_a_firmar, archivo_certificado=bytes_certificado, extension_certificado=extension_certificado,
                            password_certificado=contrasenaCertificado,
                            page=int(membrete["numPage"]), reason=razon, lx=membrete["x"], ly=membrete["y"]
                        ).sign_and_get_content_bytes()
                        documento_a_firmar = io.BytesIO()
                        documento_a_firmar.write(datau)
                        documento_a_firmar.seek(0)


                    file_obj = DjangoFile(documento_a_firmar, name=f"{_name}.pdf")
                    acta = ActaConstatacion(constatacion=constatacion,
                                            archivo=file_obj,
                                            estado=2,
                                            cantidad=len(jsonFirmas),
                                            persona=persona)
                    acta.save(request)
                    log(u'Firmo Documento: {}'.format(_name), request, "add")
                    if constatacion.total_firmas() >= 2:
                        constatacion.estadoacta = 3
                        constatacion.save(request)
                        log(u'Remito acta a usuario para firmar: {}'.format(constatacion), request, "edit")
                        titulo='Acta de constatación de activos fijos pendiente de firmar'
                        mensaje='Tiene un acta de constatación de activos que requiere ser firmada por su persona'
                        notify(titulo, mensaje, constatacion.usuariobienes, None,
                                     f'/mis_activos',
                                     constatacion.pk, 1, 'sga-sagest', ConstatacionFisica, request)
                    else:
                        constatacion.estadoacta = 2
                        constatacion.save(request)
                    return JsonResponse({'result': False, 'mensaje': 'Guardado correctamente'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, "mensaje": 'Error: {}'.format(ex)}, safe=False)

        elif action == 'firmarconstatacionmasivo':
            try:
                # FIRMA ELECTRÓNICA
                certificado = request.FILES["firma"]
                contrasenaCertificado = request.POST['palabraclave']
                razon = request.POST['razon'] if 'razon' in request.POST else ''
                extension_certificado = os.path.splitext(certificado.name)[1][1:]
                bytes_certificado = certificado.read()
                periodo = PeriodoConstatacionAF.objects.get(id=encrypt_id(request.POST['id']))
                directoraf = SeccionDepartamento.objects.get(id=23).responsable
                actas_firmadas = ActaConstatacion.objects.filter(status=True, constatacion__periodo=periodo, persona=persona, estado=2).values_list('constatacion_id', flat=True).order_by('constatacion_id').distinct()
                es_director = persona == directoraf
                if es_director:
                    constataciones = ConstatacionFisica.objects.filter(status=True, estado=2, periodo=periodo, estadoacta=2).exclude(id__in=actas_firmadas)
                else:
                    constataciones = ConstatacionFisica.objects.filter(status=True, usuariofinaliza=persona, estado=2, periodo=periodo, estadoacta=1).exclude(id__in=actas_firmadas)

                for c in constataciones:
                    if not c.acta_persona_firmada(persona):
                        try:
                            archivo_ = c.get_documento()
                            archivo_url = archivo_.url
                            # director = director_af(informe)
                            # cargo = persona.mi_cargo_administrativo() if persona.mi_cargo_administrativo() else ''
                            firmas = []
                            nombre_persona =persona.nombre_completo_minus()
                            palabras = f'{nombre_persona} Responsable de constatación'
                            x, y, numPage = obtener_posicion_x_y_saltolinea(archivo_url, palabras, True, True, True)
                            if x and y:
                                firmas.append({'x': x, 'y': y, 'numPage': numPage})
                            if es_director:
                                palabras = f'{nombre_persona} Experto de activos fijos'
                                x, y, numPage = obtener_posicion_x_y_saltolinea(archivo_url, palabras, True, True, True)
                                if x and y:
                                    firmas.append({'x': x, 'y': y, 'numPage': numPage})
                            if firmas:
                                for membrete in firmas:
                                    datau = JavaFirmaEc(
                                        archivo_a_firmar=archivo_, archivo_certificado=bytes_certificado, extension_certificado=extension_certificado,
                                        password_certificado=contrasenaCertificado,
                                        page=int(membrete["numPage"]), reason=razon, lx=membrete["x"], ly=membrete["y"]
                                    ).sign_and_get_content_bytes()
                                    archivo_ = io.BytesIO()
                                    archivo_.write(datau)
                                    archivo_.seek(0)
                                _name = generar_nombre(f"acta_constatacion_{c.usuariobienes.usuario.username}_{c.id}",'')
                                file_obj = DjangoFile(archivo_, name=f"{_name}.pdf")
                                acta = ActaConstatacion(constatacion=c,
                                                        archivo=file_obj,
                                                        estado=2,
                                                        cantidad=len(firmas),
                                                        persona=persona)
                                acta.save(request)
                                if c.total_firmas() >= 2:
                                    c.estadoacta = 3
                                    c.save(request)
                                    log(u'Remito acta a usuario para firmar: {}'.format(c), request, "edit")
                                    titulo = 'Acta de constatación de activos fijos pendiente de firmar'
                                    mensaje = 'Tiene un acta de constatación de activos que requiere ser firmada por su persona'
                                    notify(titulo, mensaje, c.usuariobienes, None,
                                           f'/mis_activos',
                                           c.pk, 1, 'sga-sagest', ConstatacionFisica, request)
                                else:
                                    c.estadoacta = 2
                                    c.save(request)
                                log(u'Guardo acta de constatación firmada: {}'.format(acta), request, "add")
                        except Exception as ex:
                            pass
                return JsonResponse({"result": False, "mensaje": "Guardado con exito", }, safe=False)
            except Exception as ex:
                print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
                textoerror = '{} Linea:{}'.format(ex, sys.exc_info()[-1].tb_lineno)
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error: %s" % ex})

        elif action == 'notificarmasivofirmacustodio':
            try:
                id = request.POST['id']
                periodo = PeriodoConstatacionAF.objects.get(id=encrypt_id(id))
                constataciones = ConstatacionFisica.objects.filter(status=True, periodo=periodo, estadoacta=3).order_by('estado', '-numero')
                titulo = 'Acta de constatación de activos fijos pendiente de firmar'
                mensaje = 'Tiene un acta de constatación de activos que requiere ser firmada por su persona'
                for c in constataciones:
                    notify(titulo, mensaje, c.usuariobienes, None,
                           f'/mis_activos',
                           c.pk, 1, 'sga-sagest', ConstatacionFisica, request)
                return JsonResponse({'result': 'ok', 'mensaje': 'Notificaciones enviadas correctamente', 'showSwal': 'true'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': "bad", "mensaje": 'Error: {}'.format(ex)}, safe=False)

        elif action == 'resetearinformebaja':
            try:
                id = int(encrypt(request.POST['id']))
                documento = DocumentoFirmaInformeBaja.objects.get(id=id)
                informebaja = documento.informe
                director = get_directorresponsablebaja('AF')
                file_obj, response = generar_informebaja_pdf_v2(informebaja, director)

                documento.director = director
                documento.informe = informebaja
                documento.estadofirma = 1
                documento.firmadirector = False
                documento.archivo = file_obj
                documento.save(request)

                historial = HistorialDocumentoInformeBaja(
                    documentoinforme=documento,
                    persona=informebaja.responsable,
                    archivo=documento.archivo,
                    estadofirma=documento.estadofirma
                )
                historial.save(request)

                activo = informebaja.activofijo
                if activo.archivobaja:
                    activo.archivobaja = None
                    activo.save(request)

                return JsonResponse({"result": 'ok', "mensaje": "Reseteado con exito", })
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": str(ex)})

        elif action == 'descargarinformestecnicos':
            try:
                noti = Notificacion(cuerpo='Se esta procesando la descarga de los informes técnicos.',
                                    titulo='Archivo .zip de reportes técnicos en proceso',
                                    destinatario=persona,
                                    url='/notificacion',
                                    prioridad=1, app_label='sga-sagest',
                                    fecha_hora_visible=datetime.now() + timedelta(days=1), tipo=2,
                                    en_proceso=True)
                noti.save(request)
                descargar_informes_baja_masivo_background(request=request, data=data, notif=noti.pk).start()
                return JsonResponse({'result': 'ok', 'mensaje': 'El archivo .zip se esta generando. Verifique su apartado de notificaciones después de unos minutos.'})
            except Exception as ex:
                return JsonResponse({'result': 'bad', 'mensaje': f'Error al generar archivo .zip {ex}'})

        elif action == 'descargaractasconstatacion':
            try:
                estado = request.POST.get('estado', '0')
                periodo = PeriodoConstatacionAF.objects.get(id=int(encrypt_id(request.POST['id'])))
                filtro = Q(status=True, periodo=periodo)
                if estado != '0':
                    filtro = filtro & Q(estadoacta=int(estado))
                if periodo:
                    constataciones = ConstatacionFisica.objects.filter(filtro)
                    if not constataciones:
                        raise NameError('No existen actas de constatación para el estado seleccionado.')
                else:
                    raise NameError('No se ha seleccionado un periodo de constatación.')
                noti = Notificacion(cuerpo='Se esta procesando la descarga de actas de constatación.',
                                    titulo='Archivo .zip de actas de constatación en proceso',
                                    destinatario=persona,
                                    url='/notificacion',
                                    prioridad=1, app_label='sga-sagest',
                                    fecha_hora_visible=datetime.now() + timedelta(days=1), tipo=2,
                                    en_proceso=True)
                noti.save(request)
                descargar_actas_constatacion_masivo_background(request=request, data=data, notif=noti.pk).start()
                return JsonResponse({'result': False, 'mensaje': 'El archivo .zip se esta generando. Verifique su apartado de notificaciones después de unos minutos.'})
            except Exception as ex:
                return JsonResponse({'result': True, 'mensaje': f'Error al generar archivo .zip {ex}'})

        return JsonResponse({"result": "bad", "mensaje": u"Solicitud Incorrecta."})
    else:
        if 'action' in request.GET:
            data['action'] = action = request.GET['action']

            if action == 'estadobien':
                try:
                    data['title'] = u'Estado de Bienes Inmuebles de la Institución'
                    data['estados'] = EstadoBien.objects.filter(status=True)
                    return render(request, "af_activofijo/view_estadobien.html", data)
                except:
                    pass

            elif action == 'addestadobien':
                try:
                    data['title'] = u'Adicionar Estado de Bienes Inmuebles de la Institución'
                    data['form'] = EstadoBienForm()
                    return render(request, "af_activofijo/addestdobien.html", data)
                except:
                    pass

            elif action == 'editestadobien':
                try:
                    data['title'] = u'Editar Estado de Bienes Inmuebles de la Institución'
                    data['estado'] = estado = EstadoBien.objects.get(pk=int(request.GET['id']))
                    data['form'] = EstadoBienForm(initial={'nombre': estado.nombre, 'descripcion': estado.descripcion})
                    return render(request, "af_activofijo/editestadobien.html", data)
                except:
                    pass

            elif action == 'delestadobien':
                try:
                    data['title'] = u'Eliminar Estado de Bienes Inmuebles de la Institución'
                    data['estado'] = EstadoBien.objects.get(pk=int(request.GET['id']))
                    return render(request, "af_activofijo/delestadobien.html", data)
                except:
                    pass

            elif action == 'condicionbien':
                try:
                    data['title'] = u'Condiciones de Bienes Inmuebles de la Institución'
                    data['condiciones'] = CondicionBien.objects.filter(status=True)
                    return render(request, "af_activofijo/view_condicionbien.html", data)
                except:
                    pass

            elif action == 'addcondicionbien':
                try:
                    data['title'] = u'Adicionar Condición de Bienes Inmuebles de la Institución'
                    data['form'] = CondicionBienForm()
                    return render(request, "af_activofijo/addcondicionbien.html", data)
                except:
                    pass

            elif action == 'editcondicionbien':
                try:
                    data['title'] = u'Editar Condición de Bienes Inmuebles de la Institución'
                    data['condicion'] = estado = CondicionBien.objects.get(pk=int(request.GET['id']))
                    data['form'] = CondicionBienForm(
                        initial={'nombre': estado.nombre, 'descripcion': estado.descripcion})
                    return render(request, "af_activofijo/editcondicionbien.html", data)
                except:
                    pass

            elif action == 'delcondicionbien':
                try:
                    data['title'] = u'Eliminar Condición de Bienes Inmuebles de la Institución'
                    data['condicion'] = CondicionBien.objects.get(pk=int(request.GET['id']))
                    return render(request, "af_activofijo/delcondicionbien.html", data)
                except:
                    pass

            elif action == 'detalle_activo':
                try:
                    data['activo'] = activo = ActivoFijo.objects.get(pk=int(request.GET['id']))
                    template = get_template("af_activofijo/detalle.html")
                    # json_content = template.render(data)
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

            elif action == 'edificios':
                try:
                    data['title'] = u'Bienes Inmuebles de la Institución'
                    # data['edificios'] = Edificio.objects.filter(status=True)
                    search = None
                    ids = None
                    if 's' in request.GET:
                        search = request.GET['s'].strip()
                        ss = search.split(' ')
                        if len(ss) == 1:
                            edificios = Edificio.objects.filter(Q(identificacion__descripcion__icontains=search) |
                                                                Q(codigobien=search) |
                                                                Q(codigoanterior=search) |
                                                                Q(catalogo__descripcion__icontains=search) |
                                                                Q(responsable__nombres__icontains=search) |
                                                                Q(responsable__apellido1__icontains=search) |
                                                                Q(responsable__apellido2__icontains=search) |
                                                                Q(cuentacontable__cuenta__icontains=search) |
                                                                Q(cuentacontable__descripcion__icontains=search) |
                                                                Q(identificador=search) |
                                                                Q(condicionbien__nombre__icontains=search) |
                                                                Q(estadobien__nombre__icontains=search)).distinct().order_by(
                                'fechaingreso')
                        else:
                            edificios = Edificio.objects.filter(Q(responsable__apellido1__icontains=ss[0]) | Q(
                                responsable__apellido2__icontains=ss[1]) |
                                                                (Q(identificacion__descripcion__icontains=ss[0]) & Q(
                                                                    identificacion__descripcion__icontains=ss[
                                                                        1]))).distinct().order_by('fechaingreso')
                    elif 'id' in request.GET:
                        ids = int(request.GET['id'])
                        edificios = Edificio.objects.filter(id=ids)
                    else:
                        edificios = Edificio.objects.filter(status=True).order_by('fechaingreso')
                    paging = MiPaginador(edificios, 15)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['edificios'] = page.object_list
                    data['reporte_0'] = obtener_reporte('edificio_cuenta')
                    data['cuentascontables'] = CuentaContable.objects.filter(status=True)
                    data['form001'] = ReporteEdificioFrom()
                    return render(request, 'af_activofijo/view_edificio.html', data)
                except Exception as ex:
                    pass

            elif action == 'mantenimientos':
                try:
                    data['title'] = u'Mantenimientos preventivos sin garantía'
                    search = None
                    ids = None
                    data['tipos'] = GruposCategoria.objects.all()
                    if 's' in request.GET:
                        search = request.GET['s'].strip()
                        ss = search.split(' ')
                        if len(ss) == 1:
                            mantenimientosactivos = MantenimientosActivosPreventivos.objects.filter(
                                Q(activofijo__codigointerno=search) | Q(activofijo__codigogobierno=search),
                                status=True).order_by('fecha')
                        else:
                            mantenimientosactivos = MantenimientosActivosPreventivos.objects.filter(
                                status=True).order_by('-fecha')
                    else:
                        # mantenimientosactivos = MantenimientosActivosPreventivos.objects.filter(status=True).order_by('-fecha')
                        mantenimientosactivos = MantenimientosActivosPreventivos.objects.filter(
                            tipoactivo__descripcion='COMPUTADOR DE ESCRITORIO').order_by('-fecha')
                    if 'tipo' in request.GET:
                        tipo = request.GET['tipo']
                        data['idtipo'] = int(tipo)
                        mantenimientosactivos = MantenimientosActivosPreventivos.objects.filter(
                            tipoactivo__id=tipo).order_by('-fecha')
                    if 'id' in request.GET:
                        ids = request.GET['id']
                        mantenimientosactivos = MantenimientosActivosPreventivos.objects.filter(id=ids).order_by(
                            '-fecha')
                    if ('fini' in request.GET) and ('ffin' in request.GET) and ('tipo' in request.GET):
                        fini = request.GET['fini']
                        ffin = request.GET['ffin']
                        tipo = request.GET['tipo']
                        mantenimientosactivos = MantenimientosActivosPreventivos.objects.filter(tipoactivo__id=tipo,
                                                                                                fecha__range=[fini,
                                                                                                              ffin])

                    paging = MiPaginador(mantenimientosactivos, 15)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['totalact'] = mantenimientosactivos.count()
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['mantenimientosactivos'] = page.object_list
                    return render(request, 'af_activofijo/view_mantenimientos.html', data)
                except Exception as ex:
                    pass

            elif action == 'tareasmantenimiento':
                try:
                    data['title'] = u'Tareas de mantenimientos'
                    search = None
                    ids = None
                    if 's' in request.GET:
                        search = request.GET['s'].strip()
                        ss = search.split(' ')
                        if len(ss) == 1:
                            tareasmantenimiento = MantenimientoGruCategoria.objects.filter(
                                descripcion__icontains=search, status=True).order_by('grupocategoria_id', 'descripcion')
                        else:
                            tareasmantenimiento = MantenimientoGruCategoria.objects.filter(
                                descripcion__icontains=search, status=True).order_by('grupocategoria_id', 'descripcion')
                    else:
                        tareasmantenimiento = MantenimientoGruCategoria.objects.filter(status=True).order_by(
                            'grupocategoria_id', 'descripcion')
                    paging = MiPaginador(tareasmantenimiento, 15)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['tareasmantenimiento'] = page.object_list
                    return render(request, 'af_activofijo/view_tareasmantenimientos.html', data)
                except Exception as ex:
                    pass

            elif action == 'tareasmantenimientosgdanio':
                try:
                    data['title'] = u'Lista de daños encontrados'
                    search = None
                    ids = None
                    if 's' in request.GET:
                        search = request.GET['s'].strip()
                        ss = search.split(' ')
                        if len(ss) == 1:
                            tareasmantenimiento = MantenimientoGruDanios.objects.filter(
                                Q(descripcion__icontains=search) | Q(grupocategoria__descripcion__icontains=search),
                                status=True).order_by('grupocategoria_id', 'descripcion')
                        else:
                            tareasmantenimiento = MantenimientoGruDanios.objects.filter(descripcion__icontains=search,
                                                                                        status=True).order_by(
                                'grupocategoria_id', 'descripcion')
                    else:
                        tareasmantenimiento = MantenimientoGruDanios.objects.filter(status=True).order_by(
                            'grupocategoria_id', 'descripcion')
                    paging = MiPaginador(tareasmantenimiento, 15)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['tareasmantenimiento'] = page.object_list
                    return render(request, 'af_activofijo/view_tareasmantenimientossgdanio.html', data)
                except Exception as ex:
                    pass

            elif action == 'addtareamantenimientosgdanios':
                try:
                    # puede_realizar_accion(request, 'sagest.puede_ingresar_compras')
                    data['title'] = u'Adicionar daño encontrado'
                    data['form'] = TareaMantenimientoDaniosForm()
                    plantilla = render(request, 'af_activofijo/addtareamantenimientosgdanios.html', data)
                    return plantilla
                except Exception as ex:
                    pass

            elif action == 'edittareamantenimientosgdanio':
                try:
                    data['title'] = u'Editar daño encontrado'
                    data['tarea'] = tarea = MantenimientoGruDanios.objects.get(status=True, pk=int(request.GET['id']))
                    data['form'] = TareaMantenimientoDaniosForm(
                        initial={'categoria': tarea.grupocategoria, 'descripcion': tarea.descripcion})
                    return render(request, "af_activofijo/edittareamantenimientosgdanio.html", data)
                except:
                    pass

            elif action == 'deltaremantenimientosgdanio':
                try:
                    data['title'] = u'Eliminar daño encontrado'
                    data['tareamantenimiento'] = MantenimientoGruDanios.objects.get(pk=int(request.GET['id']))
                    return render(request, "af_activofijo/deltareamantenimientosgdanio.html", data)
                except:
                    pass

            elif action == 'tareasmantenimientolimpieza':
                try:
                    data['title'] = u'Lista de tareas en limpieza'
                    search = None
                    ids = None
                    if 's' in request.GET:
                        search = request.GET['s'].strip()
                        ss = search.split(' ')
                        if len(ss) == 1:
                            tareaslimpieza = MantenimientoGruCategoriaGarantiaLimp.objects.filter(
                                descripcion__icontains=search, status=True).order_by('grupocategoria_id', 'descripcion')
                        else:
                            tareaslimpieza = MantenimientoGruCategoriaGarantiaLimp.objects.filter(
                                descripcion__icontains=search, status=True).order_by('grupocategoria_id', 'descripcion')
                    else:
                        tareaslimpieza = MantenimientoGruCategoriaGarantiaLimp.objects.filter(status=True).order_by(
                            'grupocategoria_id', 'descripcion')
                    paging = MiPaginador(tareaslimpieza, 15)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['tareasmantenimiento'] = page.object_list
                    return render(request, 'af_activofijo/view_tareasmantenimientoslimpieza.html', data)
                except Exception as ex:
                    pass

            elif action == 'addtareamantenimientolimpieza':
                try:
                    # puede_realizar_accion(request, 'sagest.puede_ingresar_compras')
                    data['title'] = u'Adicionar tarea de mantenimiento en limpieza'
                    data['form'] = TareaMantenimientoFrom()
                    plantilla = render(request, 'af_activofijo/addtareamantenimientolimpieza.html', data)
                    return plantilla
                except Exception as ex:
                    pass

            elif action == 'deltaremantenimientolimpieza':
                try:
                    data['title'] = u'Eliminar Tarea Mantenimiento en Limpieza'
                    data['tareamantenimiento'] = MantenimientoGruCategoriaGarantiaLimp.objects.get(
                        pk=int(request.GET['id']))
                    return render(request, "af_activofijo/deltareamantenimientolimpieza.html", data)
                except:
                    pass

            elif action == 'edittaremantenimientolimpieza':
                try:
                    data['title'] = u'Editar tarea de mantenimiento'
                    data['tarea'] = tarea = MantenimientoGruCategoriaGarantiaLimp.objects.get(status=True, pk=int(
                        encrypt(request.GET['id'])))
                    data['form'] = TareaMantenimientoFrom(
                        initial={'categoria': tarea.grupocategoria, 'descripcion': tarea.descripcion})
                    return render(request, "af_activofijo/edittareamantenimientolimpieza.html", data)
                except:
                    pass

            elif action == 'tareasmantenimientodanio':
                try:
                    data['title'] = u'Lista de daños encontrados'
                    search = None
                    ids = None
                    if 's' in request.GET:
                        search = request.GET['s'].strip()
                        ss = search.split(' ')
                        if len(ss) == 1:
                            tareasdanio = MantenimientoGruCategoriaGarantiaErr.objects.filter(
                                descripcion__icontains=search, status=True).order_by('grupocategoria_id', 'descripcion')
                        else:
                            tareasdanio = MantenimientoGruCategoriaGarantiaErr.objects.filter(
                                descripcion__icontains=search, status=True).order_by('grupocategoria_id', 'descripcion')
                    else:
                        tareasdanio = MantenimientoGruCategoriaGarantiaErr.objects.filter(status=True).order_by(
                            'grupocategoria_id', 'descripcion')
                    paging = MiPaginador(tareasdanio, 15)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['tareasmantenimiento'] = page.object_list
                    return render(request, 'af_activofijo/view_tareasmantenimientosdanio.html', data)
                except Exception as ex:
                    pass

            elif action == 'addtareamantenimientodanio':
                try:
                    # puede_realizar_accion(request, 'sagest.puede_ingresar_compras')
                    data['title'] = u'Adicionar tarea de mantenimiento en daños'
                    data['form'] = TareaMantenimientoFrom()
                    plantilla = render(request, 'af_activofijo/addtareamantenimientodanio.html', data)
                    return plantilla
                except Exception as ex:
                    pass

            elif action == 'deltaremantenimientodanio':
                try:
                    data['title'] = u'Eliminar Tarea Mantenimiento en Daños'
                    data['tareamantenimiento'] = MantenimientoGruCategoriaGarantiaErr.objects.get(
                        pk=int(request.GET['id']))
                    return render(request, "af_activofijo/deltareamantenimientodanio.html", data)
                except:
                    pass

            elif action == 'edittaremantenimientodanio':
                try:
                    data['title'] = u'Editar tarea de mantenimiento'
                    data['tarea'] = tarea = MantenimientoGruCategoriaGarantiaErr.objects.get(status=True, pk=int(
                        encrypt(request.GET['id'])))
                    data['form'] = TareaMantenimientoFrom(
                        initial={'categoria': tarea.grupocategoria, 'descripcion': tarea.descripcion})
                    return render(request, "af_activofijo/edittareamantenimientodanio.html", data)
                except:
                    pass

            elif action == 'garantiamantenimiento':
                try:
                    data['title'] = u'Mantenimientos preventivos con garantía'
                    data['tipos'] = GruposCategoria.objects.all()
                    search = None
                    ids = None
                    garantiamantenimiento = None
                    if 's' in request.GET:
                        search = request.GET['s'].strip()
                        if search.isdigit():
                            garantiamantenimiento = MantenimientosActivosGarantia.objects.filter(
                                Q(activofijo__codigogobierno=search) | Q(activofijo__codigointerno=search) | Q(
                                    detallemantenimientosactivosgarantia__activofijo__codigogobierno=search) | Q(
                                    detallemantenimientosactivosgarantia__activofijo__codigointerno=search),
                                status=True)
                        else:
                            ss = search.split(' ')
                            if len(ss) == 1:
                                garantiamantenimiento = MantenimientosActivosGarantia.objects.filter(
                                    Q(activofijo__descripcion__icontains=search) | Q(
                                        detallemantenimientosactivosgarantia__activofijo__descripcion__icontains=search),
                                    status=True)
                            else:
                                garantiamantenimiento = MantenimientosActivosGarantia.objects.filter(
                                    Q(activofijo__descripcion__icontains=search) | Q(
                                        detallemantenimientosactivosgarantia__activofijo__descripcion__icontains=search),
                                    status=True)
                    else:
                        if 'id' in request.GET:
                            ids = request.GET['id']
                            garantiamantenimiento = MantenimientosActivosGarantia.objects.filter(pk=ids, status=True)
                        else:
                            if 'hist' in request.GET:
                                if request.GET['hist'] == '1':
                                    garantiamantenimiento = MantenimientosActivosGarantia.objects.filter(
                                        tipoactivo__descripcion='COMPUTADOR DE ESCRITORIO', status=True)
                                    data['histbus'] = '1'
                                else:
                                    garantiamantenimiento = MantenimientosActivosGarantia.objects.filter(status=True)
                                    data['histbus'] = '2'
                    if 'tipo' in request.GET:
                        tipo = request.GET['tipo']
                        data['idtipo'] = int(tipo)
                        data['histbus'] = '1'
                        garantiamantenimiento = MantenimientosActivosGarantia.objects.filter(tipoactivo__id=tipo,
                                                                                             status=True)
                    if ('fini' in request.GET) and ('ffin' in request.GET):
                        fini = request.GET['fini']
                        ffin = request.GET['ffin']
                        if 'hist' in request.GET:
                            hist = request.GET['hist']
                            if hist == '1':
                                tipo = request.GET['tipo']
                                garantiamantenimiento = MantenimientosActivosGarantia.objects.filter(
                                    tipoactivo__id=tipo, fechainicio__range=[fini, ffin], status=True)
                            elif hist == '2':
                                garantiamantenimiento = MantenimientosActivosGarantia.objects.filter(
                                    fechainicio__range=[fini, ffin], status=True)
                    paging = MiPaginador(garantiamantenimiento, 15)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['page'] = page
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['totalact'] = garantiamantenimiento.count()
                    data['garantiamantenimiento'] = page.object_list
                    return render(request, 'af_activofijo/garantiamantenimiento.html', data)
                except Exception as ex:
                    pass

            elif action == 'catalogo':
                try:
                    data['title'] = u'Catálogo Bienes'
                    search = None
                    ids = None
                    if 's' in request.GET:
                        search = request.GET['s'].strip()
                        ss = search.split(' ')
                        if len(ss) == 1:
                            catalogos = CatalogoBien.objects.filter(
                                Q(descripcion__icontains=search)).distinct().order_by('fechaingreso')
                        else:
                            catalogos = Edificio.objects.filter(Q(responsable__apellido1__icontains=ss[0]) | Q(
                                responsable__apellido2__icontains=ss[1]) |
                                                                (Q(identificacion__descripcion__icontains=ss[0]) & Q(
                                                                    identificacion__descripcion__icontains=ss[
                                                                        1]))).distinct().order_by('fechaingreso')
                    elif 'id' in request.GET:
                        ids = int(request.GET['id'])
                        catalogos = CatalogoBien.objects.filter(id=ids)
                    else:
                        catalogos = CatalogoBien.objects.filter(status=True).order_by('descripcion')
                    paging = MiPaginador(catalogos, 15)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['catalogos'] = page.object_list
                    return render(request, 'af_activofijo/view_catalogo.html', data)
                except Exception as ex:
                    pass

            elif action == 'inventariotecnologico':
                try:
                    data['title'] = u'Listado'
                    return render(request, 'af_activofijo/listadoinventario.html', data)
                except Exception as ex:
                    pass

            elif action == 'clasificar':
                try:
                    data['title'] = u'Listado de categorías por clasificar'
                    data['catalogos'] = CatalogoBien.objects.filter(status=True, clasificado=False)

                    return render(request, 'af_activofijo/clasificar.html', data)
                except Exception as ex:
                    pass

            elif action == 'categorizar':
                try:
                    data['title'] = u'Listado para agrupar por categorías'
                    data['catalogos'] = CatalogoBien.objects.filter(status=True, clasificado=True,
                                                                    equipoelectronico=True)

                    return render(request, 'af_activofijo/categorizar.html', data)
                except Exception as ex:
                    pass

            elif action == 'chartinventariotecnologico':
                try:
                    data['title'] = u'Equipo Tecnológico'
                    data['codigo'] = codigo = int(request.GET['codigo'])
                    data['rangosemaforo'] = anios = RangoVidaUtil.objects.filter(status=True).order_by('anio',
                                                                                                       'descripcion')
                    data['grupocatalogo'] = GruposCategoria.objects.filter(status=True)
                    if codigo == 0:
                        data['listadocatalogo'] = activos = ActivoFijo.objects.filter(
                            Q(archivobaja__isnull=True) | Q(archivobaja=''), catalogo__equipoelectronico=True,
                            catalogo__status=True, statusactivo=1, status=True).order_by('descripcion')
                    else:
                        data['listadocatalogo'] = activos = ActivoFijo.objects.filter(
                            Q(archivobaja__isnull=True) | Q(archivobaja=''), catalogo__equipoelectronico=True,
                            catalogo__status=True, catalogo__grupo__id=codigo, statusactivo=1, status=True).order_by(
                            'descripcion')
                    data['totales'] = activos.values('id').count()
                    return render(request, "af_activofijo/charinventariotecnologico.html", data)
                except Exception as ex:
                    pass

            elif action == 'addedificio':
                try:
                    # puede_realizar_accion(request, 'sagest.puede_ingresar_compras')
                    data['title'] = u'Adicionar Bien Inmueble de la Institución'
                    data['form'] = EdificioForm()
                    plantilla = render(request, 'af_activofijo/addedificio.html', data)
                    return plantilla
                except Exception as ex:
                    pass

            elif action == 'addtareamantenimiento':
                try:
                    # puede_realizar_accion(request, 'sagest.puede_ingresar_compras')
                    data['title'] = u'Adicionar tarea de mantenimiento'
                    data['form'] = TareaMantenimientoFrom()
                    plantilla = render(request, 'af_activofijo/addtareamantenimiento.html', data)
                    return plantilla
                except Exception as ex:
                    pass

            elif action == 'edittaremantenimiento':
                try:
                    data['title'] = u'Editar tarea de mantenimiento'
                    data['tarea'] = tarea = MantenimientoGruCategoria.objects.get(status=True,
                                                                                  pk=int(request.GET['id']))
                    data['form'] = TareaMantenimientoFrom(
                        initial={'categoria': tarea.grupocategoria, 'descripcion': tarea.descripcion})
                    return render(request, "af_activofijo/edittareamantenimiento.html", data)
                except:
                    pass

            elif action == 'addmantenimientogarantia':
                try:
                    # puede_realizar_accion(request, 'sagest.puede_ingresar_compras')
                    data['title'] = u'Nuevo Mantenimiento Equipo Tecnológico Con Garantía'
                    data['form'] = MantenimientosActivosGarantiaForm()
                    data['form3'] = ActivosFijosForm()
                    data['grupocategoria'] = GruposCategoria.objects.filter(status=True)
                    plantilla = render(request, 'af_activofijo/addmantenimientogarantia.html', data)
                    return plantilla
                except Exception as ex:
                    pass

            elif action == 'editedificio':
                try:
                    data['title'] = u'Editar Bienes Inmuebles de la Institución'
                    data['edificio'] = edificio = Edificio.objects.get(pk=int(request.GET['id']))
                    initial = model_to_dict(edificio)
                    form = EdificioForm(initial=initial)
                    form.fields['catalogo'].widget.attrs['descripcion'] = edificio.catalogo
                    form.fields['catalogo'].widget.attrs['value'] = edificio.catalogo.id
                    form.fields['responsable'].widget.attrs['descripcion'] = edificio.responsable
                    form.fields['responsable'].widget.attrs['value'] = edificio.responsable.id
                    form.fields['custodio'].widget.attrs['descripcion'] = edificio.custodio
                    form.fields['custodio'].widget.attrs['value'] = edificio.custodio.id
                    data['form'] = form
                    return render(request, "af_activofijo/editedificio.html", data)
                except:
                    pass

            elif action == 'editmantenimientogarantia':
                try:
                    ingnuevo = request.GET['nuevo']
                    data['title'] = u'Editar Mantenimiento Garantía'
                    data['mangarantia'] = mangarantia = MantenimientosActivosGarantia.objects.get(
                        pk=int(encrypt(request.GET['idgarantia'])))
                    #
                    data['grupocategoria'] = GruposCategoria.objects.filter(status=True)
                    # Llamando a los componentes de daños y limpieza
                    data['tareaslimpieza'] = MantenimientoGruCategoriaGarantiaLimp.objects.filter(
                        grupocategoria=mangarantia.tipoactivo, status=True)
                    data['tareasdanio'] = MantenimientoGruCategoriaGarantiaErr.objects.filter(
                        grupocategoria=mangarantia.tipoactivo, status=True)
                    data['tareaslimpiezat'] = TareasActivosPreventivosGarantiaLimp.objects.values_list('grupos_id',
                                                                                                       flat=True).filter(
                        mantenimiento=mangarantia, status=True)
                    data['tareasdaniot'] = TareasActivosPreventivosGarantiaErr.objects.values_list('grupos_id',
                                                                                                   flat=True).filter(
                        mantenimiento=mangarantia, status=True)
                    #
                    data['detallemantenimiento'] = mangarantia.detallemantenimientosactivosgarantia_set.filter(
                        status=True).order_by('activofijo__descripcion')
                    form = MantenimientosActivosGarantiaForm(initial={'numreporte': mangarantia.numreporte,
                                                                      'fechainicio': mangarantia.fechainicio,
                                                                      'valor': mangarantia.valor,
                                                                      'tipoactivo': mangarantia.tipoactivo,
                                                                      'observacion': mangarantia.observacion,
                                                                      'horamax': mangarantia.horamax,
                                                                      'minutomax': mangarantia.minutomax,
                                                                      'estfrec': mangarantia.estfrec,
                                                                      'estfent': mangarantia.estfent})
                    data['form'] = form
                    # from sga.forms import ReservasCraiSolicitarAutoridadForm
                    data['form3'] = ActivosFijosForm()
                    if (ingnuevo == 'False'):
                        return render(request, "af_activofijo/editmantenimientogarantia.html", data)
                    else:
                        return render(request, "af_activofijo/editmantenimientogarantianuevo.html", data)
                except Exception as ex:
                    pass

            elif action == 'deledificio':
                try:
                    data['title'] = u'Eliminar Bienes Inmuebles de la Institución'
                    data['edificio'] = Edificio.objects.get(pk=int(request.GET['id']))
                    return render(request, "af_activofijo/deledificio.html", data)
                except:
                    pass

            elif action == 'delmantenimiento':
                try:
                    data['title'] = u'Eliminar Mantenimiento'
                    data['mantenimiento'] = MantenimientosActivosPreventivos.objects.get(pk=int(request.GET['id']))
                    return render(request, "af_activofijo/delmantenimiento.html", data)
                except:
                    pass

            elif action == 'delmantenimientogarantia':
                try:
                    data['title'] = u'Eliminar Mantenimiento Garantía'
                    data['mantenimientogarantia'] = MantenimientosActivosGarantia.objects.get(pk=int(request.GET['id']))
                    return render(request, "af_activofijo/delmantenimientogarantia.html", data)
                except:
                    pass

            elif action == 'deltaremantenimiento':
                try:
                    data['title'] = u'Eliminar Tarea Mantenimiento'
                    data['tareamantenimiento'] = MantenimientoGruCategoria.objects.get(pk=int(request.GET['id']))
                    return render(request, "af_activofijo/deltareamantenimiento.html", data)
                except:
                    pass

            elif action == 'buscarresponsable_custodio':
                try:
                    q = request.GET['q'].upper().strip()
                    s = q.split(" ")
                    if s.__len__() == 2:
                        persona = Persona.objects.filter(status=True, administrativo__isnull=False,
                                                         nombres__contains=s[0], apellido1__icontains=s[0],
                                                         apellido2__icontains=s[0]).distinct()[:20]
                    else:
                        persona = Persona.objects.filter((Q(nombres__contains=s[0]) | Q(apellido1__contains=s[0]) | Q(
                            apellido2__contains=s[0]) | Q(cedula__contains=s[0]))).filter(status=True,
                                                                                          administrativo__isnull=False).distinct()[
                                  :20]
                    data = {"result": "ok", "results": [{"id": x.id, "name": x.flexbox_repr()} for x in persona]}
                    return JsonResponse((data))
                except Exception as ex:
                    pass

            elif action == 'add':
                try:
                    # puede_realizar_accion(request, 'sagest.puede_ingresar_compras')
                    data['title'] = u'Nuevo Activo'
                    form = ActivoFijoForm()
                    data['form'] = form
                    plantilla = render(request, 'af_activofijo/add.html', data)
                    return plantilla
                except Exception as ex:
                    pass

            elif action == 'addmantenimiento':
                try:
                    # puede_realizar_accion(request, 'sagest.puede_ingresar_compras')
                    data['title'] = u'Nuevo Mantenimiento Equipo Tecnológico Sin Garantía'
                    # data['tiposmantenimiento'] = MantenimientoGruCategoria.objects.filter(status=True)
                    data['form'] = MantenimientosActivosPreventivosForm()
                    data['grupocategoria'] = GruposCategoria.objects.filter(status=True)
                    data['form3'] = ActivosFijosForm()
                    # data['tipo'] = TIPO_MANTENIMIENTO
                    return render(request, "af_activofijo/addmantenimiento.html", data)
                except Exception as ex:
                    pass

            elif action == 'editmantenimiento':
                try:
                    # puede_realizar_accion(request, 'sagest.puede_ingresar_compras')
                    data['title'] = u'Editar Mantenimiento'
                    data['manpreventivos'] = mantenimiento = MantenimientosActivosPreventivos.objects.get(
                        pk=request.GET['idman'])
                    data['grupocategoria'] = GruposCategoria.objects.filter(status=True)
                    data['tareasmantenimiento'] = MantenimientoGruCategoria.objects.filter(
                        grupocategoria=mantenimiento.tipoactivo, status=True, activo=True)
                    data['tareasactivo'] = TareasActivosPreventivos.objects.values_list('grupos_id', flat=True).filter(
                        mantenimiento=mantenimiento, status=True)
                    data['piezaparteactivo'] = PiezaParteActivosPreventivos.objects.filter(mantenimiento=mantenimiento,
                                                                                           status=True)
                    data['daniomantenimiento'] = MantenimientoGruDanios.objects.filter(
                        grupocategoria=mantenimiento.tipoactivo, status=True, activo=True)
                    data['danioactivo'] = TareasActivosPreventivosDanios.objects.filter(
                        mantenimiento=mantenimiento, status=True)
                    data['tipo'] = TIPO_MANTENIMIENTO
                    form = MantenimientosActivosPreventivosForm(initial={'tipoactivo': mantenimiento.tipoactivo,
                                                                         'estusu': mantenimiento.estusu,
                                                                         'archivo': mantenimiento.archivo,
                                                                         'fecha': mantenimiento.fecha,
                                                                         'horamax': mantenimiento.horamax,
                                                                         'minutomax': mantenimiento.minutomax,
                                                                         'estfrec': mantenimiento.funcionarecibe,
                                                                         'estfent': mantenimiento.funcionaentrega,
                                                                         'marca': mantenimiento.marca,
                                                                         'modelo': mantenimiento.modelo,
                                                                         # 'caracteristica': mantenimiento.caracteristicas,
                                                                         'bsugiere': mantenimiento.sbequipo,
                                                                         'dsugiere': mantenimiento.descbaja,
                                                                         # 'piezaparte': piezaparte,
                                                                         'observacion': mantenimiento.observaciones})
                    form.cargar_mantenimiento(mantenimiento)
                    # cat = HdPiezaPartes.objects.filter(pk=mantenimiento.tipoactivo)
                    # form.fields['piezaparte'].queryset = cat
                    data['form'] = form
                    data['form3'] = ActivosFijosForm()
                    if mantenimiento.nuevo:
                        return render(request, "af_activofijo/editmantenimientov2.html", data)
                    else:
                        return render(request, "af_activofijo/editmantenimiento.html", data)
                except Exception as ex:
                    pass

            elif action == 'edit':
                try:
                    data['title'] = u'Modificar Activo'
                    data['activo'] = activo = ActivoFijo.objects.get(pk=request.GET['id'])
                    form = ActivoFijoForm(initial={
                        'codigogobierno': activo.codigogobierno,
                        'codigointerno': activo.codigointerno,
                        'estructuraactivo': activo.estructuraactivo,
                        'clasebien': activo.clasebien,
                        'fechaingreso': activo.fechaingreso,
                        'descripcion': activo.descripcion,
                        'observacion': activo.observacion,
                        'origeningreso': activo.origeningreso,
                        'costo': activo.costo,
                        'serie': activo.serie,
                        'modelo': activo.modelo,
                        'marca': activo.marca,
                        'tipodocumentorespaldo': activo.tipodocumentorespaldo,
                        'clasedocumentorespaldo': activo.clasedocumentorespaldo,
                        'numerocomprobante': activo.numerocomprobante,
                        'tipocomprobante': activo.tipocomprobante,
                        'fechacomprobante': activo.fechacomprobante,
                        'estado': activo.estado,
                        'tipoproyecto': activo.tipoproyecto,
                        'deprecia': activo.deprecia,
                        'vidautil': activo.vidautil,
                        'cuentacontable': activo.cuentacontable,
                        'ubicacion': activo.ubicacion,
                        'titulo': activo.titulo,
                        'autor': activo.autor,
                        'editorial': activo.editorial,
                        'fechaedicion': activo.fechaedicion,
                        'numeroedicion': activo.numeroedicion,
                        'clasificacionbibliografica': activo.clasificacionbibliografica,
                        'color': activo.color,
                        'material': activo.material,
                        'dimensiones': activo.dimensiones,
                        'clasevehiculo': activo.clasevehiculo,
                        'tipovehiculo': activo.tipovehiculo,
                        'numeromotor': activo.numeromotor,
                        'numerochasis': activo.numerochasis,
                        'placa': activo.placa,
                        'aniofabricacion': activo.aniofabricacion,
                        'colorprimario': activo.colorprimario,
                        'colorsecundario': activo.colorsecundario,
                        'propietario': activo.propietario,
                        'codigocatastral': activo.codigocatastral,
                        'numeropredio': activo.numeropredio,
                        'valoravaluo': activo.valoravaluo,
                        'anioavaluo': activo.anioavaluo,
                        'areapredio': activo.areapredio,
                        'areaconstruccion': activo.areaconstruccion,
                        'pisos': activo.pisos,
                        'provincia': activo.provincia,
                        'canton': activo.canton,
                        'parroquia': activo.parroquia,
                        'zona': activo.zona,
                        'nomenclatura': activo.nomenclatura,
                        'sector': activo.sector,
                        'direccion': activo.direccion,
                        'direccion2': activo.direccion2,
                        'escritura': activo.escritura,
                        'fechaescritura': activo.fechaescritura,
                        'notaria': activo.notaria,
                        'beneficiariocontrato': activo.beneficiariocontrato,
                        'fechacontrato': activo.fechacontrato,
                        'duracioncontrato': activo.duracioncontrato,
                        'montocontrato': activo.montocontrato,
                        'catalogo': activo.catalogo.id,
                        'fechainiciogarantia': activo.fechainiciogarantia,
                        'fechafingarantia': activo.fechafingarantia,
                        'ebye': activo.ebye,
                    })
                    form.editar(activo, request.user)
                    data['form'] = form
                    return render(request, 'af_activofijo/edit.html', data)
                except Exception as ex:
                    pass

            elif action == 'gestionarbaja':
                try:
                    activo = ActivoFijo.objects.get(pk=int(encrypt(request.GET['id'])))
                    data['id'] = activo.id
                    data['activo'] = activo
                    # data['codigogobierno'] = activo.codigogobierno
                    # data['descripcion'] = activo.descripcion
                    form = GestionarBajaActivoForm()
                    data['form'] = form
                    data['puedegestionarbaja'] = activo.estado.id == 3 or (activo.estado.id == 2 and activo.condicionestado == 2)
                    template = get_template('af_activofijo/modal/formgestionarbaja.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'addingresoinformebaja':
                try:
                    data['title'] = u'Ingreso informe de baja'
                    data['activofijo'] = activofijo = ActivoFijo.objects.get(pk=int(encrypt(request.GET['idactivo'])))
                    data['tipo'] = tipo = (request.GET['tipo'])
                    data['formdetalle'] = DetalleInformeBajaForm()
                    if activofijo.responsable:
                        idsolicita = activofijo.responsable.id
                    else:
                        idsolicita = 0
                    data['codsolicita'] = idsolicita
                    if tipo == 1:
                        form = InformeBajaForm(initial={'solicita': activofijo.responsable,
                                                        'estadouso': 2,
                                                        'estado': 2})
                    else:
                        form = InformeBajaFormAF(initial={'solicita': activofijo.responsable,
                                                         'estadouso': 2,
                                                         'estado': 2})
                    data['form'] = form
                    return render(request, 'af_activofijo/addingresobaja.html', data)
                except Exception as ex:
                    pass

            elif action == 'editingresoinformebaja':
                try:
                    data['title'] = u'Modificar Activo'
                    data['activoinformebaja'] = activoinformebaja = InformeActivoBaja.objects.get(
                        activofijo_id=int(encrypt(request.GET['idactivo'])), status=True)
                    if activoinformebaja.tipoinforme == 1:
                        form = InformeBajaForm(initial={
                            'solicita': activoinformebaja.solicita,
                            'responsable': activoinformebaja.responsable,
                            'conclusion': activoinformebaja.conclusion,
                            'estado': activoinformebaja.estado,
                            'estadouso': activoinformebaja.estadouso,
                            'bloque': activoinformebaja.bloque,
                            'detallerevision': activoinformebaja.detallerevision})
                    else:
                        form = InformeBajaFormAF(initial={
                            'solicita': activoinformebaja.solicita,
                            'responsable': activoinformebaja.responsable,
                            'conclusion': activoinformebaja.conclusion,
                            'estado': activoinformebaja.estado,
                            'estadouso': activoinformebaja.estadouso,
                            'bloque': activoinformebaja.bloque,
                            'detallerevision': activoinformebaja.detallerevision,
                            'departamento': activoinformebaja.departamento,
                            'gestion': activoinformebaja.gestion})

                    data['form'] = form
                    data['formdetalle'] = DetalleInformeBajaForm()
                    data['detalleinformebaja'] = activoinformebaja.detalleinformeactivobaja_set.filter(status=True)
                    return render(request, 'af_activofijo/editingresobaja.html', data)
                except Exception as ex:
                    pass

            elif action == 'excellistadoactivos':
                try:
                    fdesde, fhasta = request.GET.get('fechadesde', ''), request.GET.get('fechahasta', '')
                    __author__ = 'Unemi'
                    style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
                    style_nb = easyxf('font: name Times New Roman, color-index blue, bold on',
                                      num_format_str='#,##0.00')
                    style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
                    title = easyxf(
                        'font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                    style1 = easyxf(num_format_str='D-MMM-YY')
                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('exp_xls_post_part')
                    # ws.write_merge(0, 0, 0, 10, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                    response = HttpResponse(content_type="application/ms-excel")
                    response[
                        'Content-Disposition'] = 'attachment; filename=Listado' + random.randint(1,
                                                                                                 10000).__str__() + '.xls'
                    columns = [
                        (u"RESPONSABLE", 10000),
                        (u"CODIGO INTERNO", 2000),
                        (u"CPDOGP GOBIERNO", 2000),
                        (u"FECHA INGRESO", 4000),
                        (u"TIEMPO", 4500),
                        (u"NUM. AÑOS", 2000),
                        (u"CATALOGO", 15000),
                        (u"DESCRIPCIÓN", 10000),
                        (u"MODELO", 10000),
                        (u"MARCA", 10000),
                        (u"DOCUMENTO BAJA", 10000),
                        (u"FECHA BAJA", 10000),
                        (u"UBICACIÓN", 10000),
                        (u"ESTADO", 10000),
                        (u"ARCHIVO DE BAJA", 10000),
                    ]
                    row_num = 0
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], font_style)
                        ws.col(col_num).width = columns[col_num][1]
                    date_format = xlwt.XFStyle()
                    date_format.num_format_str = 'yyyy/mm/dd'
                    lista1 = '0'
                    if 'cadenatexto' in request.GET:
                        lista1 = request.GET['cadenatexto']
                    if lista1 == '0' and not fdesde and not fhasta:
                        sql = "select ac.codigointerno,ac.codigogobierno,ac.fechaingreso,age (current_date, ac.fechaingreso) || '' tiempo, " \
                              "cast(extract(year from age (current_date, ac.fechaingreso))*12 + extract(month from age (ac.fechaingreso)) as int)/12 as numanios " \
                              ",cat.descripcion as catalogo,ac.descripcion,ac.modelo,ac.marca," \
                              "(select perso.apellido1 || ' ' || perso.apellido2 || ' ' || perso.nombres from sga_persona perso where id=ac.responsable_id) as responsable , " \
                              "ac.id, (SELECT ubi.nombre AS ubicacion FROM sagest_ubicacion ubi WHERE ubi.id = ac.ubicacion_id)," \
                              "(SELECT est.nombre AS estado FROM sagest_estadoproducto est WHERE est.id = ac.estado_id), ac.archivobaja " \
                              "from sagest_activofijo ac,sagest_catalogobien cat,sagest_gruposcategoria gru " \
                              "where ac.catalogo_id=cat.id  and cat.equipoelectronico=True and cat.grupo_id=gru.id " \
                              " and ac.status=True and (ac.archivobaja='' or ac.archivobaja isnull)"
                    if fdesde and fhasta and lista1 == '0':
                        sql = "select ac.codigointerno,ac.codigogobierno,ac.fechaingreso,age (current_date, ac.fechaingreso) || '' tiempo, " \
                              "cast(extract(year from age (current_date, ac.fechaingreso))*12 + extract(month from age (ac.fechaingreso)) as int)/12 as numanios " \
                              ",cat.descripcion as catalogo,ac.descripcion,ac.modelo,ac.marca," \
                              "(select perso.apellido1 || ' ' || perso.apellido2 || ' ' || perso.nombres from sga_persona perso where id=ac.responsable_id) as responsable , " \
                              "ac.id, (SELECT ubi.nombre AS ubicacion FROM sagest_ubicacion ubi WHERE ubi.id = ac.ubicacion_id)," \
                              "(SELECT est.nombre AS estado FROM sagest_estadoproducto est WHERE est.id = ac.estado_id), ac.archivobaja " \
                              "from sagest_activofijo ac,sagest_catalogobien cat,sagest_gruposcategoria gru " \
                              "where ac.catalogo_id=cat.id  and cat.equipoelectronico=True and cat.grupo_id=gru.id " \
                              " and ac.status=True and (ac.archivobaja='' or ac.archivobaja isnull) AND ac.fechaingreso >= '{}' AND ac.fechaingreso <= '{}'".format(
                            fdesde, fhasta)
                    if lista1 != '0' and not fdesde and not fhasta:
                        sql = "select ac.codigointerno,ac.codigogobierno,ac.fechaingreso,age (current_date, ac.fechaingreso) || '' tiempo, " \
                              "cast(extract(year from age (current_date, ac.fechaingreso))*12 + extract(month from age (ac.fechaingreso)) as int)/12 as numanios " \
                              ",cat.descripcion as catalogo,ac.descripcion,ac.modelo,ac.marca,(select perso.apellido1 || ' ' || perso.apellido2 || ' ' || perso.nombres from sga_persona perso where id=ac.responsable_id) as responsable ," \
                              " ac.id, (SELECT ubi.nombre AS ubicacion FROM sagest_ubicacion ubi WHERE ubi.id = ac.ubicacion_id), " \
                              " (SELECT est.nombre AS estado FROM sagest_estadoproducto est WHERE est.id = ac.estado_id), ac.archivobaja " \
                              "from sagest_activofijo ac,sagest_catalogobien cat,sagest_gruposcategoria gru " \
                              "where ac.catalogo_id=cat.id  and cat.equipoelectronico=True and cat.grupo_id=gru.id " \
                              " and ac.status=True and gru.id=" + lista1 + " and (ac.archivobaja='' or ac.archivobaja isnull)"
                    if lista1 != '0' and fdesde and fhasta:
                        sql = "select ac.codigointerno,ac.codigogobierno,ac.fechaingreso,age (current_date, ac.fechaingreso) || '' tiempo, " \
                              "cast(extract(year from age (current_date, ac.fechaingreso))*12 + extract(month from age (ac.fechaingreso)) as int)/12 as numanios " \
                              ",cat.descripcion as catalogo,ac.descripcion,ac.modelo,ac.marca,(select perso.apellido1 || ' ' || perso.apellido2 || ' ' || perso.nombres from sga_persona perso where id=ac.responsable_id) as responsable ," \
                              " ac.id, (SELECT ubi.nombre AS ubicacion FROM sagest_ubicacion ubi WHERE ubi.id = ac.ubicacion_id), " \
                              " (SELECT est.nombre AS estado FROM sagest_estadoproducto est WHERE est.id = ac.estado_id), ac.archivobaja " \
                              "from sagest_activofijo ac,sagest_catalogobien cat,sagest_gruposcategoria gru " \
                              "where ac.catalogo_id=cat.id  and cat.equipoelectronico=True and cat.grupo_id=gru.id " \
                              " and ac.status=True and gru.id=" + lista1 + " and (ac.archivobaja='' or ac.archivobaja isnull) AND ac.fechaingreso >= '{}' AND ac.fechaingreso <= '{}'".format(
                            fdesde, fhasta)
                    cursor = connection.cursor()
                    cursor.execute(sql)
                    results = cursor.fetchall()
                    row_num = 1
                    for r in results:
                        campo1 = r[0]
                        campo2 = r[1]
                        campo3 = r[2]
                        campo4 = r[3]
                        campo5 = r[4]
                        campo6 = r[5]
                        campo7 = r[6]
                        campo8 = r[7]
                        campo9 = r[8]
                        campo10 = r[9]
                        campo11 = 'NO'
                        campo12 = ''
                        if InformeActivoBaja.objects.filter(status=True, activofijo_id=r[10]).exists():
                            informe = InformeActivoBaja.objects.get(status=True, activofijo_id=r[10])
                            campo11 = 'SI'
                            campo12 = str(informe.fecha_creacion)
                        campo13 = r[11]
                        campo14 = r[12]
                        campo15 = r[13]
                        ws.write(row_num, 0, campo10, font_style2)
                        ws.write(row_num, 1, campo1, font_style2)
                        ws.write(row_num, 2, campo2, font_style2)
                        ws.write(row_num, 3, campo3, date_format)
                        ws.write(row_num, 4, campo4, font_style2)
                        ws.write(row_num, 5, campo5, font_style2)
                        ws.write(row_num, 6, campo6, font_style2)
                        ws.write(row_num, 7, campo7, font_style2)
                        ws.write(row_num, 8, campo8, font_style2)
                        ws.write(row_num, 9, campo9, font_style2)
                        ws.write(row_num, 10, campo11, font_style2)
                        ws.write(row_num, 11, campo12, font_style2)
                        ws.write(row_num, 12, campo13, font_style2)
                        ws.write(row_num, 13, campo14, font_style2)
                        ws.write(row_num, 14, "https://sagest.unemi.edu.ec/media/".format(campo15), font_style2)
                        row_num += 1
                    wb.save(response)
                    return response
                except Exception as ex:
                    pass

            elif action == 'pdflistadoactivos':
                try:
                    data['title'] = nombrearchivo = u'Informe de Activos Tecnológicos'
                    fdesde, fhasta = request.GET.get('fechadesde', ''), request.GET.get('fechahasta', '')
                    data['desde'] = str(fdesde)
                    data['hasta'] = str(fhasta)
                    lista1 = '0'
                    if 'cadenatexto' in request.GET:
                        lista1 = request.GET['cadenatexto']
                    values_ac = InformeActivoBaja.objects.values('activofijo__id').filter(status=True).values_list(
                        'activofijo__id', flat=True)
                    if lista1 == '0':
                        sql = "select ac.codigointerno,ac.codigogobierno,ac.fechaingreso,age (current_date, ac.fechaingreso) || '' tiempo, " \
                              "cast(extract(year from age (current_date, ac.fechaingreso))*12 + extract(month from age (ac.fechaingreso)) as int)/12 as numanios " \
                              ",cat.descripcion as catalogo,ac.descripcion,ac.modelo,ac.marca," \
                              "(select perso.apellido1 || ' ' || perso.apellido2 || ' ' || perso.nombres from sga_persona perso where id=ac.responsable_id) as responsable , " \
                              "ac.id, (SELECT ubi.nombre AS ubicacion FROM sagest_ubicacion ubi WHERE ubi.id = ac.ubicacion_id)," \
                              "(SELECT est.nombre AS estado FROM sagest_estadoproducto est WHERE est.id = ac.estado_id) " \
                              "from sagest_activofijo ac,sagest_catalogobien cat,sagest_gruposcategoria gru " \
                              "where ac.catalogo_id=cat.id  and cat.equipoelectronico=True and cat.grupo_id=gru.id " \
                              "and ac.statusactivo=1 and ac.status=True and (ac.archivobaja='' or ac.archivobaja isnull) " \
                              "AND ac.fechaingreso >= '{}' AND ac.fechaingreso <= '{}' AND NOT ac.id IN {} ORDER BY ac.fechaingreso ASC".format(
                            fdesde, fhasta, str(tuple(values_ac)))
                    else:
                        sql = "select ac.codigointerno,ac.codigogobierno,ac.fechaingreso,age (current_date, ac.fechaingreso) || '' tiempo, " \
                              "cast(extract(year from age (current_date, ac.fechaingreso))*12 + extract(month from age (ac.fechaingreso)) as int)/12 as numanios " \
                              ",cat.descripcion as catalogo,ac.descripcion,ac.modelo,ac.marca," \
                              "(select perso.apellido1 || ' ' || perso.apellido2 || ' ' || perso.nombres from sga_persona perso where id=ac.responsable_id) as responsable , " \
                              "ac.id, (SELECT ubi.nombre AS ubicacion FROM sagest_ubicacion ubi WHERE ubi.id = ac.ubicacion_id)," \
                              "(SELECT est.nombre AS estado FROM sagest_estadoproducto est WHERE est.id = ac.estado_id) " \
                              "from sagest_activofijo ac,sagest_catalogobien cat,sagest_gruposcategoria gru " \
                              "where ac.catalogo_id=cat.id  and cat.equipoelectronico=True and cat.grupo_id=gru.id " \
                              "and ac.statusactivo=1 and ac.status=True and gru.id=" + lista1 + " and (ac.archivobaja='' or ac.archivobaja isnull) " \
                                                                                                "AND ac.fechaingreso >= '{}' AND ac.fechaingreso <= '{}' AND NOT ac.id IN {} ORDER BY ac.fechaingreso ASC".format(
                            fdesde, fhasta, str(tuple(values_ac)))
                    cursor = connection.cursor()
                    cursor.execute(sql)
                    results = cursor.fetchall()
                    data['results'] = results
                    data['hoy'] = str(datetime.now().date())
                    # return conviert_html_to_pdf_name(
                    #     'af_activofijo/informes/activotecnologicopdf.html',
                    #     {'pagesize': 'A4', 'data': data,}, nombrearchivo)
                    return conviert_html_to_pdf(
                        'af_activofijo/informes/activotecnologicopdf.html',
                        {
                            'pagesize': 'A4',
                            'data': data,
                        },
                    )
                except Exception as ex:
                    pass

            elif action == 'pdflistadoactivosinactios':
                try:
                    data['title'] = nombrearchivo = u'Informe de Activos Tecnológicos'
                    fdesde, fhasta = request.GET.get('fechadesde', ''), request.GET.get('fechahasta', '')
                    data['desde'] = str(fdesde)
                    data['hasta'] = str(fhasta)
                    lista1 = '0'
                    if 'cadenatexto' in request.GET:
                        lista1 = request.GET['cadenatexto']
                    values_ac = InformeActivoBaja.objects.values('activofijo__id').filter(status=True,
                                                                                          activofijo__catalogo__equipoelectronico=True,
                                                                                          fecha_creacion__gte=fdesde,
                                                                                          fecha_creacion__lte=fhasta).values_list(
                        'activofijo__id', flat=True)
                    if not values_ac:
                        values_ac = [0]
                    if lista1 == '0':
                        sql = "select ac.codigointerno,ac.codigogobierno,ac.fechaingreso,age (current_date, ac.fechaingreso) || '' tiempo, " \
                              "cast(extract(year from age (current_date, ac.fechaingreso))*12 + extract(month from age (ac.fechaingreso)) as int)/12 as numanios " \
                              ",cat.descripcion as catalogo,ac.descripcion,ac.modelo,ac.marca," \
                              "(select perso.apellido1 || ' ' || perso.apellido2 || ' ' || perso.nombres from sga_persona perso where id=ac.responsable_id) as responsable , " \
                              "ac.id, (SELECT ubi.nombre AS ubicacion FROM sagest_ubicacion ubi WHERE ubi.id = ac.ubicacion_id)," \
                              "(SELECT est.nombre AS estado FROM sagest_estadoproducto est WHERE est.id = ac.estado_id), " \
                              "(SELECT DATE(fecha_creacion) FROM sagest_informeactivobaja WHERE activofijo_id=ac.id) AS fecha_salida " \
                              "from sagest_activofijo ac,sagest_catalogobien cat,sagest_gruposcategoria gru " \
                              "where ac.catalogo_id=cat.id  and cat.equipoelectronico=True and cat.grupo_id=gru.id " \
                              "and ac.statusactivo=1 and ac.status=True and (ac.archivobaja='' or ac.archivobaja isnull) " \
                              "AND ac.id IN {} ORDER BY fecha_salida ASC".format(str(tuple(values_ac)))
                    else:
                        sql = "select ac.codigointerno,ac.codigogobierno,ac.fechaingreso,age (current_date, ac.fechaingreso) || '' tiempo, " \
                              "cast(extract(year from age (current_date, ac.fechaingreso))*12 + extract(month from age (ac.fechaingreso)) as int)/12 as numanios " \
                              ",cat.descripcion as catalogo,ac.descripcion,ac.modelo,ac.marca," \
                              "(select perso.apellido1 || ' ' || perso.apellido2 || ' ' || perso.nombres from sga_persona perso where id=ac.responsable_id) as responsable , " \
                              "ac.id, (SELECT ubi.nombre AS ubicacion FROM sagest_ubicacion ubi WHERE ubi.id = ac.ubicacion_id)," \
                              "(SELECT est.nombre AS estado FROM sagest_estadoproducto est WHERE est.id = ac.estado_id), " \
                              "(SELECT DATE(fecha_creacion) FROM sagest_informeactivobaja WHERE activofijo_id=ac.id) AS fecha_salida " \
                              "from sagest_activofijo ac,sagest_catalogobien cat,sagest_gruposcategoria gru " \
                              "where ac.catalogo_id=cat.id  and cat.equipoelectronico=True and cat.grupo_id=gru.id " \
                              "and ac.statusactivo=1 and ac.status=True and gru.id=" + lista1 + " and (ac.archivobaja='' or ac.archivobaja isnull) " \
                                                                                                "AND  ac.id IN {} ORDER BY fecha_salida ASC".format(
                            str(tuple(values_ac)))
                    cursor = connection.cursor()
                    cursor.execute(sql)
                    results = cursor.fetchall()
                    data['results'] = results
                    data['hoy'] = str(datetime.now().date())
                    # return conviert_html_to_pdf_name(
                    #     'af_activofijo/informes/activotecnologicopdf.html',
                    #     {'pagesize': 'A4', 'data': data,}, nombrearchivo)
                    return conviert_html_to_pdf(
                        'af_activofijo/informes/bajasactivotecnologicopdf.html',
                        {
                            'pagesize': 'A4',
                            'data': data,
                        },
                    )
                except Exception as ex:
                    pass

            elif action == 'pdflistadoactivostodos':
                try:
                    data['title'] = nombrearchivo = u'Informe de Activos Tecnológicos'
                    lista1 = '0'
                    if 'cadenatexto' in request.GET:
                        lista1 = request.GET['cadenatexto']
                    values_ac = InformeActivoBaja.objects.values('activofijo__id').filter(status=True).values_list(
                        'activofijo__id', flat=True)
                    if lista1 == '0':
                        sql = "select ac.codigointerno,ac.codigogobierno,ac.fechaingreso,age (current_date, ac.fechaingreso) || '' tiempo, " \
                              "cast(extract(year from age (current_date, ac.fechaingreso))*12 + extract(month from age (ac.fechaingreso)) as int)/12 as numanios " \
                              ",cat.descripcion as catalogo,ac.descripcion,ac.modelo,ac.marca," \
                              "(select perso.apellido1 || ' ' || perso.apellido2 || ' ' || perso.nombres from sga_persona perso where id=ac.responsable_id) as responsable , " \
                              "ac.id, (SELECT ubi.nombre AS ubicacion FROM sagest_ubicacion ubi WHERE ubi.id = ac.ubicacion_id)," \
                              "(SELECT est.nombre AS estado FROM sagest_estadoproducto est WHERE est.id = ac.estado_id) " \
                              "from sagest_activofijo ac,sagest_catalogobien cat,sagest_gruposcategoria gru " \
                              "where ac.catalogo_id=cat.id  and cat.equipoelectronico=True and cat.grupo_id=gru.id " \
                              "and ac.statusactivo=1 and ac.status=True and (ac.archivobaja='' or ac.archivobaja isnull) " \
                              "AND NOT ac.id IN {} ORDER BY ac.fechaingreso ASC".format(str(tuple(values_ac)))
                    else:
                        sql = "select ac.codigointerno,ac.codigogobierno,ac.fechaingreso,age (current_date, ac.fechaingreso) || '' tiempo, " \
                              "cast(extract(year from age (current_date, ac.fechaingreso))*12 + extract(month from age (ac.fechaingreso)) as int)/12 as numanios " \
                              ",cat.descripcion as catalogo,ac.descripcion,ac.modelo,ac.marca," \
                              "(select perso.apellido1 || ' ' || perso.apellido2 || ' ' || perso.nombres from sga_persona perso where id=ac.responsable_id) as responsable , " \
                              "ac.id, (SELECT ubi.nombre AS ubicacion FROM sagest_ubicacion ubi WHERE ubi.id = ac.ubicacion_id)," \
                              "(SELECT est.nombre AS estado FROM sagest_estadoproducto est WHERE est.id = ac.estado_id) " \
                              "from sagest_activofijo ac,sagest_catalogobien cat,sagest_gruposcategoria gru " \
                              "where ac.catalogo_id=cat.id  and cat.equipoelectronico=True and cat.grupo_id=gru.id " \
                              "and ac.statusactivo=1 and ac.status=True and gru.id=" + lista1 + " and (ac.archivobaja='' or ac.archivobaja isnull) " \
                                                                                                "AND NOT ac.id IN {} ORDER BY ac.fechaingreso ASC".format(
                            str(tuple(values_ac)))
                    cursor = connection.cursor()
                    cursor.execute(sql)
                    results = cursor.fetchall()
                    data['results'] = results
                    data['hoy'] = str(datetime.now().date())
                    # return conviert_html_to_pdf_name(
                    #     'af_activofijo/informes/activotecnologicopdf.html',
                    #     {'pagesize': 'A4', 'data': data,}, nombrearchivo)
                    return conviert_html_to_pdf(
                        'af_activofijo/informes/activotecnologicopdf.html',
                        {
                            'pagesize': 'A4',
                            'data': data,
                        },
                    )
                except Exception as ex:
                    pass

            elif action == 'pdflistadoactivosinactiostodos':
                try:
                    data['title'] = nombrearchivo = u'Informe de Activos Tecnológicos'
                    fdesde, fhasta = request.GET.get('fechadesde', ''), request.GET.get('fechahasta', '')
                    data['desde'] = str(fdesde)
                    data['hasta'] = str(fhasta)
                    lista1 = '0'
                    if 'cadenatexto' in request.GET:
                        lista1 = request.GET['cadenatexto']
                    values_ac = InformeActivoBaja.objects.values('activofijo__id').filter(status=True,
                                                                                          activofijo__catalogo__equipoelectronico=True,
                                                                                          fecha_creacion__gte=fdesde,
                                                                                          fecha_creacion__lte=fhasta).values_list(
                        'activofijo__id', flat=True)
                    if not values_ac:
                        values_ac = [0]
                    if lista1 == '0':
                        sql = "select ac.codigointerno,ac.codigogobierno,ac.fechaingreso,age (current_date, ac.fechaingreso) || '' tiempo, " \
                              "cast(extract(year from age (current_date, ac.fechaingreso))*12 + extract(month from age (ac.fechaingreso)) as int)/12 as numanios " \
                              ",cat.descripcion as catalogo,ac.descripcion,ac.modelo,ac.marca," \
                              "(select perso.apellido1 || ' ' || perso.apellido2 || ' ' || perso.nombres from sga_persona perso where id=ac.responsable_id) as responsable , " \
                              "ac.id, (SELECT ubi.nombre AS ubicacion FROM sagest_ubicacion ubi WHERE ubi.id = ac.ubicacion_id)," \
                              "(SELECT est.nombre AS estado FROM sagest_estadoproducto est WHERE est.id = ac.estado_id), " \
                              "(SELECT DATE(fecha_creacion) FROM sagest_informeactivobaja WHERE activofijo_id=ac.id) AS fecha_salida " \
                              "from sagest_activofijo ac,sagest_catalogobien cat,sagest_gruposcategoria gru " \
                              "where ac.catalogo_id=cat.id  and cat.equipoelectronico=True and cat.grupo_id=gru.id " \
                              "and ac.statusactivo=1 and ac.status=True and (ac.archivobaja='' or ac.archivobaja isnull) " \
                              "AND ac.id IN {} ORDER BY fecha_salida ASC".format(str(tuple(values_ac)))
                    else:
                        sql = "select ac.codigointerno,ac.codigogobierno,ac.fechaingreso,age (current_date, ac.fechaingreso) || '' tiempo, " \
                              "cast(extract(year from age (current_date, ac.fechaingreso))*12 + extract(month from age (ac.fechaingreso)) as int)/12 as numanios " \
                              ",cat.descripcion as catalogo,ac.descripcion,ac.modelo,ac.marca," \
                              "(select perso.apellido1 || ' ' || perso.apellido2 || ' ' || perso.nombres from sga_persona perso where id=ac.responsable_id) as responsable , " \
                              "ac.id, (SELECT ubi.nombre AS ubicacion FROM sagest_ubicacion ubi WHERE ubi.id = ac.ubicacion_id)," \
                              "(SELECT est.nombre AS estado FROM sagest_estadoproducto est WHERE est.id = ac.estado_id), " \
                              "(SELECT DATE(fecha_creacion) FROM sagest_informeactivobaja WHERE activofijo_id=ac.id) AS fecha_salida " \
                              "from sagest_activofijo ac,sagest_catalogobien cat,sagest_gruposcategoria gru " \
                              "where ac.catalogo_id=cat.id  and cat.equipoelectronico=True and cat.grupo_id=gru.id " \
                              "and ac.statusactivo=1 and ac.status=True and gru.id=" + lista1 + " and (ac.archivobaja='' or ac.archivobaja isnull) " \
                                                                                                "AND  ac.id IN {} ORDER BY fecha_salida ASC".format(
                            str(tuple(values_ac)))
                    cursor = connection.cursor()
                    cursor.execute(sql)
                    results = cursor.fetchall()
                    data['results'] = results
                    data['hoy'] = str(datetime.now().date())
                    # return conviert_html_to_pdf_name(
                    #     'af_activofijo/informes/activotecnologicopdf.html',
                    #     {'pagesize': 'A4', 'data': data,}, nombrearchivo)
                    return conviert_html_to_pdf(
                        'af_activofijo/informes/bajasactivotecnologicopdf.html',
                        {
                            'pagesize': 'A4',
                            'data': data,
                        },
                    )
                except Exception as ex:
                    pass

            elif action == 'editacta':
                try:
                    data['title'] = u'Editar Acta'
                    data['acta'] = acta = TraspasoActivo.objects.get(pk=request.GET['id'])
                    detalle = acta.detalletraspasoactivo_set.all()[0]
                    actas = acta.detalletraspasoactivo_set.all()
                    form = ActasForm(initial={
                        'numero': acta.numero,
                        'fecha': acta.fecha_creacion,
                        'tipobien': detalle.activo.catalogo.tipobien,
                        'tipocomprobante': detalle.activo.tipocomprobante,
                        'origeningreso': detalle.activo.origeningreso,
                        'proveedor': acta.proveedor,
                        'responsable': acta.usuariobienrecibe,
                        'custodio': acta.custodiobienrecibe,
                        'ubicacion': acta.ubicacionbienrecibe,
                        'observacion': acta.observacion,
                    })
                    form.editar()
                    data['form'] = form
                    paging = MiPaginador(actas, 10)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['detalles'] = page.object_list
                    data['usuario'] = request.user
                    return render(request, 'af_activofijo/editacta.html', data)
                except Exception as ex:
                    pass

            elif action == 'consultacatalogo':
                try:
                    data['title'] = u'Consultas de activos por catálogos'
                    data['form'] = ConsultacatalogoForm()
                    data['permite_modificar'] = False
                    data['reporte_0'] = obtener_reporte('consulta_activos_catalogo')
                    return render(request, 'af_activofijo/consultacatalogo.html', data)
                except Exception as ex:
                    pass

            elif action == 'consultausuario':
                try:
                    data['title'] = u'Consultas de activos por usuario'
                    data['form'] = ConsultausuarioForm()
                    data['permite_modificar'] = False
                    data['reporte_0'] = obtener_reporte('consulta_activos_usuario')
                    return render(request, 'af_activofijo/consultausuario.html', data)
                except Exception as ex:
                    pass

            elif action == 'asignacion':
                try:
                    data['title'] = u'Asignar usuario y custodio'
                    data['traspasos'] = TraspasoActivo.objects.filter(tipo=1)
                    return render(request, "af_activofijo/asignacion.html", data)
                except Exception as ex:
                    pass

            elif action == 'asignar':
                try:
                    data['title'] = u'Asignar activos'
                    data['form'] = AsignacionActivoForm()
                    data['activos'] = ActivoFijo.objects.filter(detalletraspasoactivo__isnull=True)
                    return render(request, "af_activofijo/asignar.html", data)
                except Exception as ex:
                    pass

            elif action == 'historial':
                try:
                    data['title'] = u'Historial del Bien'
                    data['activo'] = activo = ActivoFijo.objects.get(pk=int(request.GET['id']))
                    # data["detalles"] = activo.detalletraspasoactivo_set.filter(seleccionado=True).order_by('-codigotraspaso__fecha')
                    data["detalles"] = activo.detalletraspasoactivo_set.filter(status=True).order_by(
                        '-codigotraspaso__fecha')
                    data['reporte_0'] = obtener_reporte('historial_activos')
                    data['usuario'] = request.user
                    return render(request, "af_activofijo/historial.html", data)
                except Exception as ex:
                    pass

            elif action == 'constataciones':
                try:
                    data['title'] = u'Constataciones del Bien'
                    data['activo'] = activo = ActivoFijo.objects.get(pk=int(request.GET['id']))
                    data["detalles_pert"] = activo.detalleconstatacionfisica_set.filter(perteneceusuario=True).order_by(
                        '-codigoconstatacion__numero')
                    data['detallenopert'] = activo.detalleconstatacionfisica_set.filter(perteneceusuario=False)
                    data['detallenotras'] = activo.detalleconstatacionfisica_set.filter(requieretraspaso=True)
                    data['reporte_0'] = obtener_reporte('historial_activos')
                    data['usuario'] = request.user
                    return render(request, "af_activofijo/detalle_constataciones.html", data)
                except Exception as ex:
                    pass

            elif action == 'importar':
                try:
                    data['title'] = u'Importar lista de Activos'
                    data['form'] = ImportarArchivoXLSForm()
                    return render(request, "af_activofijo/importar.html", data)
                except Exception as ex:
                    pass

            elif action == 'exportar':
                try:
                    data['title'] = u'Exportar lista de Activos'
                    data['form'] = ExportacionForm()
                    return render(request, "af_activofijo/exportar.html", data)
                except Exception as ex:
                    pass

            elif action == 'subir':
                try:
                    data['title'] = u'Subir Archivo'
                    data['importacion'] = importacion = ArchivoActivoFijo.objects.get(pk=request.GET['id'])
                    form = ImportarArchivoXLSForm()
                    form.editar()
                    data['form'] = form
                    return render(request, "af_activofijo/subir.html", data)
                except Exception as ex:
                    pass

            elif action == 'importaciones':
                try:
                    data['title'] = u'Importaciones'
                    search = None
                    ids = None
                    if 's' in request.GET:
                        search = request.GET['s']
                        importacion = ArchivoActivoFijo.objects.filter(Q(tipobien__nombre__icontains=search) |
                                                                       Q(nombre__icontains=search)).order_by('-id')
                    elif 'id' in request.GET:
                        ids = request.GET['id']
                        importacion = ArchivoActivoFijo.objects.filter(id=ids).order_by('-id')
                    else:
                        importacion = ArchivoActivoFijo.objects.all().order_by('-id')
                    paging = MiPaginador(importacion, 25)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['reporte_0'] = obtener_reporte('acta_constatacion')
                    data['importaciones'] = page.object_list
                    return render(request, "af_activofijo/importaciones.html", data)
                except Exception as ex:
                    pass

            elif action == 'exportaciones':
                try:
                    data['title'] = u'Exportaciones'
                    search = None
                    ids = None
                    if 's' in request.GET:
                        search = request.GET['s']
                        exportaciones = ExportacionesActivos.objects.filter(
                            Q(cuentacontable__descripcion__icontains=search))
                    elif 'id' in request.GET:
                        ids = request.GET['id']
                        exportaciones = ExportacionesActivos.objects.filter(id=ids)
                    else:
                        exportaciones = ExportacionesActivos.objects.all()
                    paging = MiPaginador(exportaciones, 25)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['reporte_0'] = obtener_reporte('acta_constatacion')
                    data['exportaciones'] = page.object_list
                    return render(request, "af_activofijo/exportaciones.html", data)
                except Exception as ex:
                    pass

            elif action == 'addtraspaso':
                try:
                    puede_realizar_accion(request, 'sagest.puede_modificar_traspasos')
                    data['title'] = u'Nuevo Traspaso'
                    form = TraspasoActivoForm()
                    form.adicionar()
                    data['form'] = form
                    return render(request, "af_activofijo/addtraspaso.html", data)
                except Exception as ex:
                    pass

            elif action == 'addtraspasocustodio':
                try:
                    puede_realizar_accion(request, 'sagest.puede_modificar_traspasos')
                    data['title'] = u'Nuevo Traspaso de Custodios'
                    form = TraspasoActivoCustodioForm()
                    form.adicionarcustodio()
                    data['form'] = form
                    return render(request, "af_activofijo/addtraspasocustodio.html", data)
                except Exception as ex:
                    pass

            elif action == 'addbaja':
                try:
                    puede_realizar_accion(request, 'sagest.puede_ingresar_compras')
                    data['title'] = u'Nueva Baja de Activos'
                    data['form'] = BajaActivoForm()
                    data['form2'] = DetalleBajaActivoForm()
                    return render(request, "af_activofijo/addbaja.html", data)
                except Exception as ex:
                    pass

            elif action == 'addtipobaja':
                try:
                    data['title'] = u'Nuevo Tipo Baja'
                    data['form'] = TipoBajaForm()
                    return render(request, "af_activofijo/addtipobaja.html", data)
                except Exception as ex:
                    pass

            elif action == 'addtraslado':
                try:
                    data['title'] = u'Nuevo Traslado de Activos'
                    data['form'] = TrasladoMantenimientoForm()
                    return render(request, "af_activofijo/addtraslado.html", data)
                except Exception as ex:
                    pass

            elif action == 'edittraslado':
                try:
                    data['title'] = u'Editar Traslado de Activos'
                    data['traslado'] = traslado = TrasladoMantenimiento.objects.get(pk=request.GET['id'])
                    form = TrasladoMantenimientoForm(initial={'departamentosolicita': traslado.departamentosolicita,
                                                              'asistentelogistica': traslado.asistentelogistica,
                                                              'taller': traslado.taller,
                                                              'observacion': traslado.observacion,
                                                              'administradorcontrato': traslado.administradorcontrato,
                                                              'usuariobienes': traslado.usuariobienes})
                    form.editar()
                    data['form'] = form
                    return render(request, "af_activofijo/edittraslado.html", data)
                except Exception as ex:
                    pass

            elif action == 'addconstatacion':
                try:
                    data['title'] = u'Nueva Constatación'
                    form = ConstatacionForm()
                    form.adicionar()
                    data['form'] = form
                    return render(request, "af_activofijo/addconstatacion.html", data)
                except Exception as ex:
                    pass

            elif action == 'editcons':
                try:
                    data['title'] = u'Editar Constatación'
                    data['constatacion'] = constatacion = ConstatacionFisica.objects.get(pk=request.GET['id'])
                    form = ConstatacionForm(initial={'usuariobienes': constatacion.usuariobienes.id,
                                                     'ubicacionbienes': constatacion.ubicacionbienes,
                                                     'numero': constatacion.numero,
                                                     'observacion': constatacion.observacion})
                    form.editar(constatacion)
                    data['form'] = form
                    data['form3'] = DetalleNoIdentificadoForm()
                    return render(request, "af_activofijo/editconstatacion.html", data)
                except Exception as ex:
                    pass

            elif action == 'cambiousuario':
                try:
                    data['title'] = u'Editar Constatación'
                    data['constatacion'] = constatacion = ConstatacionFisica.objects.get(pk=request.GET['id'])
                    form = ConstatacionForm(initial={'usuariobienes': constatacion.usuariobienes.id,
                                                     'ubicacionbienes': constatacion.ubicacionbienes,
                                                     'numero': constatacion.numero,
                                                     'observacion': constatacion.observacion})
                    form.cambiar(constatacion)
                    data['form'] = form
                    return render(request, "af_activofijo/cambiousuario.html", data)
                except Exception as ex:
                    pass

            elif action == 'edittraspaso':
                try:
                    puede_realizar_accion(request, 'sagest.puede_modificar_traspasos')
                    data['title'] = u'Editar Traspaso de activos'
                    data['traspaso'] = traspaso = TraspasoActivo.objects.get(pk=request.GET['id'])
                    data['detalles'] = traspaso.detalletraspasoactivo_set.all().order_by('activo__codigointerno',
                                                                                         'activo__codigogobierno')
                    initial = model_to_dict(traspaso)
                    form = TraspasoActivoForm(initial=initial)
                    form.editar()
                    data['form'] = form
                    return render(request, "af_activofijo/edittraspaso.html", data)
                except Exception as ex:
                    pass

            elif action == 'edittraspasocustodio':
                try:
                    puede_realizar_accion(request, 'sagest.puede_modificar_traspasos')
                    data['title'] = u'Editar Traspaso de activos'
                    data['traspaso'] = traspaso = TraspasoActivo.objects.get(pk=request.GET['id'])
                    data['detalles'] = traspaso.detalletraspasoactivo_set.all().order_by('activo__codigointerno',
                                                                                         'activo__codigogobierno')
                    initial = model_to_dict(traspaso)
                    form = TraspasoActivoCustodioForm(initial=initial)
                    form.editarcustodio()
                    data['form'] = form
                    return render(request, "af_activofijo/edittraspasocustodio.html", data)
                except Exception as ex:
                    pass

            elif action == 'editbaja':
                try:
                    puede_realizar_accion(request, 'sagest.puede_modificar_bajas')
                    data['title'] = u'Editar Baja de activos'
                    data['baja'] = baja = BajaActivo.objects.get(pk=request.GET['id'])
                    data['detalles'] = baja.detallebajaactivo_set.all().order_by('activo__codigointerno',
                                                                                 'activo__codigogobierno')
                    initial = model_to_dict(baja)
                    form = BajaActivoForm(initial=initial)
                    form.adicionar()
                    data['form'] = form
                    return render(request, "af_activofijo/editbaja.html", data)
                except Exception as ex:
                    pass

            elif action == 'edittipobaja':
                try:
                    data['title'] = u'Editar Tipo Baja'
                    data['tipobaja'] = tipobaja = TipoBaja.objects.get(pk=request.GET['id'])
                    initial = model_to_dict(tipobaja)
                    form = TipoBajaForm(initial=initial)
                    data['form'] = form
                    return render(request, "af_activofijo/edittipobaja.html", data)
                except Exception as ex:
                    pass

            elif action == 'histconstatacion':
                try:
                    data['title'] = u'Constatación Física'
                    search = None
                    ids = None
                    url_vars = f"&action=histconstatacion"
                    constatacion = ConstatacionFisica.objects.filter(status=True, periodo__isnull=True).order_by('estado', '-numero')
                    if 's' in request.GET:
                        search = request.GET['s']
                        ss = search.split(' ')
                        if len(ss) == 1:
                            constatacion = constatacion.filter(Q(numero__icontains=search) |
                                                               Q(usuariobienes__nombres__icontains=search) |
                                                               Q(usuariobienes__apellido1__icontains=search) |
                                                               Q(usuariobienes__apellido2__icontains=search) |
                                                               Q(ubicacionbienes__nombre__icontains=search)).distinct()
                        else:
                            constatacion = constatacion.filter(Q(usuariobienes__apellido1__icontains=ss[0]) &
                                                               Q(usuariobienes__apellido2__icontains=ss[1])).distinct()
                        url_vars += f"&s={search}"
                    elif 'id' in request.GET:
                        ids = int(request.GET['id'])
                        constatacion = constatacion.filter(id=ids)
                        url_vars += f"&id={ids}"
                    paging = MiPaginador(constatacion, 25)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['reporte_0'] = obtener_reporte('acta_constatacion')
                    data['constataciones'] = page.object_list
                    data['url_vars'] = url_vars
                    data['usuario'] = request.user
                    data['fecha'] = datetime.now().date()
                    return render(request, "af_activofijo/constataciones.html", data)
                except Exception as ex:
                    pass

            elif action == 'histconstatacionencontrados':
                try:
                    data['title'] = u'Constatación Física Bienes Encontrados'
                    search = None
                    ids = None
                    # constatacion = ConstatacionFisica.objects.filter(status=True).order_by('estado', '-numero')
                    if 's' in request.GET:
                        search = request.GET['s']
                        ss = search.split(' ')
                        if len(ss) == 1:
                            constatacion = DetalleConstatacionFisica.objects.filter(
                                Q(activo__codigogobierno__icontains=search) |
                                Q(activo__codigointerno__icontains=search), encontrado=True, enuso=True, status=True,
                                codigoconstatacion__status=True).distinct().order_by('codigoconstatacion__estado',
                                                                                     '-codigoconstatacion__numero')
                        else:
                            constatacion = DetalleConstatacionFisica.objects.filter(encontrado=True, enuso=True,
                                                                                    status=True,
                                                                                    codigoconstatacion__status=True).distinct().order_by(
                                'codigoconstatacion__estado', '-codigoconstatacion__numero')

                    # elif 'id' in request.GET:
                    #     ids = int(request.GET['id'])
                    #     constatacion = constatacion.filter(id=ids)
                    else:
                        constatacion = DetalleConstatacionFisica.objects.filter(encontrado=True, enuso=True,
                                                                                status=True,
                                                                                codigoconstatacion__status=True).distinct().order_by(
                            'codigoconstatacion__estado', '-codigoconstatacion__numero')
                    paging = MiPaginador(constatacion, 25)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['reporte_0'] = obtener_reporte('acta_constatacion')
                    data['constataciones'] = page.object_list
                    data['usuario'] = request.user
                    return render(request, "af_activofijo/constatacionesencontrados.html", data)
                except Exception as ex:
                    pass

            elif action == 'finalizarcons':
                try:
                    data['title'] = u'Confirmar finalizar constatación'
                    data['constatacion'] = ConstatacionFisica.objects.get(pk=int(request.GET['id']))
                    return render(request, "af_activofijo/finalizar.html", data)
                except:
                    pass

            elif action == 'finalizarconstatacion':
                try:
                    data['constatacion']= c =ConstatacionFisica.objects.get(pk=int(request.GET['id']))
                    data['id'] = c.id
                    data['hoy'] = hoy
                    template=get_template('af_activofijo/modal/finalizarconstatacion.html')
                    return JsonResponse({'result':True,'data':template.render(data)})
                except Exception as ex:
                    JsonResponse({'result':False,'mensaje':f'Error:{ex}'})

            elif action == 'confirmarexpo':
                try:
                    data['title'] = u'Confirmar Activos Como subidos a Gobierno'
                    data['exportacion'] = ExportacionesActivos.objects.get(pk=int(request.GET['id']))
                    return render(request, "af_activofijo/confirmaexpo.html", data)
                except:
                    pass

            elif action == 'finalizartraslado':
                try:
                    data['title'] = u'Confirmar finalizar traslado'
                    data['traslado'] = TrasladoMantenimiento.objects.get(pk=int(request.GET['id']))
                    return render(request, "af_activofijo/finalizartraslado.html", data)
                except:
                    pass

            elif action == 'finalizartraspaso':
                try:
                    data['title'] = u'Confirmar finalizar traspaso'
                    data['traspaso'] = TraspasoActivo.objects.get(pk=int(request.GET['id']))
                    return render(request, "af_activofijo/finalizartraspaso.html", data)
                except:
                    pass

            elif action == 'abrirtraspaso':
                try:
                    data['title'] = u'Confirmar reabrir traspaso'
                    data['traspaso'] = TraspasoActivo.objects.get(pk=int(request.GET['id']))
                    return render(request, "af_activofijo/abrirtraspaso.html", data)
                except:
                    pass

            elif action == 'eliminartraspaso':
                try:
                    data['title'] = u'Confirmar eliminar traspaso'
                    data['traspaso'] = TraspasoActivo.objects.get(pk=int(request.GET['id']))
                    return render(request, "af_activofijo/eliminartraspaso.html", data)
                except:
                    pass

            elif action == 'deleteexpo':
                try:
                    data['title'] = u'Confirmar eliminar exportacion'
                    data['exportacion'] = ExportacionesActivos.objects.get(pk=int(request.GET['id']))
                    return render(request, "af_activofijo/deleteexpo.html", data)
                except:
                    pass

            elif action == 'deleteimpo':
                try:
                    data['title'] = u'Confirmar eliminar importacion'
                    data['importacion'] = ArchivoActivoFijo.objects.get(pk=int(request.GET['id']))
                    return render(request, "af_activofijo/deleteimpo.html", data)
                except:
                    pass

            elif action == 'eliminartraslado':
                try:
                    data['title'] = u'Confirmar eliminar traslado mantenimiento'
                    data['traslado'] = TrasladoMantenimiento.objects.get(pk=int(request.GET['id']))
                    return render(request, "af_activofijo/eliminartraslado.html", data)
                except:
                    pass

            elif action == 'eliminarconstatacion':
                try:
                    data['title'] = u'Confirmar eliminar constatacion'
                    data['constatacion'] = ConstatacionFisica.objects.get(pk=int(request.GET['id']))
                    return render(request, "af_activofijo/eliminarconstatacion.html", data)
                except:
                    pass

            elif action == 'eliminarbaja':
                try:
                    data['title'] = u'Confirmar eliminar baja'
                    data['baja'] = BajaActivo.objects.get(pk=int(request.GET['id']))
                    return render(request, "af_activofijo/eliminarbaja.html", data)
                except:
                    pass

            elif action == 'eliminartipobaja':
                try:
                    data['title'] = u'Confirmar eliminar tipo baja'
                    data['tipobaja'] = TipoBaja.objects.get(pk=int(request.GET['id']))
                    return render(request, "af_activofijo/eliminartipobaja.html", data)
                except:
                    pass

            elif action == 'finalizarbaja':
                try:
                    data['title'] = u'Confirmar finalizar baja de activos'
                    data['baja'] = BajaActivo.objects.get(pk=int(request.GET['id']))
                    return render(request, "af_activofijo/finalizarbaja.html", data)
                except:
                    pass

            elif action == 'finalizarmantenimiento':
                try:
                    data['title'] = u'Confirmar finalizar de la mantenimiento'
                    data['mantenimiento'] = DetalleTrasladoMantenimiento.objects.get(pk=int(request.GET['id']))
                    return render(request, "af_activofijo/finalizarman.html", data)
                except:
                    pass

            elif action == 'finalizaracta':
                try:
                    data['title'] = u'Confirmar finalización de esta acta de Entrega'
                    data['acta'] = TraspasoActivo.objects.get(pk=int(request.GET['id']))
                    return render(request, "af_activofijo/finalizaracta.html", data)
                except:
                    pass

            elif action == 'histbajas':
                try:
                    data['title'] = u'Bajas de Activos'
                    search = None
                    ids = None
                    tipo=int(request.GET.get('tipo',1))
                    url_vars = f"&action=histbajas"
                    baja = BajaActivo.objects.filter(status=True).order_by('estado', '-numero')
                    if 's' in request.GET:
                        search = request.GET['s']
                        ss = search.split(' ')
                        url_vars += f"&s={search}"
                        if len(ss) == 1 and tipo == 1:
                            baja = baja.filter(Q(numero__icontains=search) |
                                               Q(solicitante__nombres__icontains=search) |
                                               Q(solicitante__apellido1__icontains=search) |
                                               Q(solicitante__apellido2__icontains=search) |
                                               Q(usuariobienentrega__nombres__icontains=search) |
                                               Q(usuariobienentrega__apellido1__icontains=search) |
                                               Q(usuariobienentrega__apellido2__icontains=search) |
                                               Q(ubicacionbienentrega__nombre__icontains=search)).distinct().order_by(
                                'estado', '-numero')
                        elif len(ss) == 1 and tipo == 2:
                            ids_a = DetalleBajaActivo.objects.filter((Q(activo__codigogobierno__icontains=search) |
                                                                      Q(activo__codigointerno__icontains=search))).values_list('codigobaja_id', flat=True).distinct()
                            baja=baja.filter(id__in=ids_a).order_by('estado', '-numero')
                            data['tipo']=tipo
                            url_vars += f"&tipo={tipo}"
                        else:
                            baja = baja.filter(Q(usuariobienentrega__apellido1__icontains=ss[0]) &
                                               Q(usuariobienentrega__apellido2__icontains=ss[1])).distinct().order_by(
                                'estado', '-numero')
                    elif 'id' in request.GET:
                        ids = request.GET['id']
                        baja = baja.filter(id=ids).order_by('estado', '-numero')
                        url_vars += f"&id={ids}"
                    paging = MiPaginador(baja, 50)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['bajas'] = page.object_list
                    data['reporte_0'] = obtener_reporte('bajas_activo')
                    data['usuario'] = request.user
                    data['url_vars'] = url_vars
                    return render(request, "af_activofijo/bajas.html", data)
                except Exception as ex:
                    pass

            elif action == 'tiposbajas':
                try:
                    data['title'] = u'Mantenimiento Tipo Bajas'
                    search = None
                    ids = None
                    tipobaja = TipoBaja.objects.filter(status=True).order_by('id')
                    if 's' in request.GET:
                        search = request.GET['s']
                        ss = search.split(' ')
                        if len(ss) == 1:
                            tipobaja = tipobaja.filter(nombre__icontains=search).distinct().order_by('id')
                        else:
                            tipobaja = tipobaja.filter(Q(nombre__icontains=ss[0]) &
                                                       Q(nombre__icontains=ss[1])).distinct().order_by('id')
                    elif 'id' in request.GET:
                        ids = request.GET['id']
                        tipobaja = tipobaja.filter(id=ids).order_by('id')
                    paging = MiPaginador(tipobaja, 25)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['tipobajas'] = page.object_list
                    data['usuario'] = request.user
                    return render(request, "af_activofijo/tiposbajas.html", data)
                except Exception as ex:
                    pass

            elif action == 'addtarjeta':
                try:
                    puede_realizar_accion(request, 'sagest.puede_modificar_tarjeta')
                    data['title'] = u'Nuevo Detalle de Mantenimiento y Reparación.'
                    data['activo'] = activo = ActivoFijo.objects.get(pk=request.GET['id'])
                    data['tarjeta'] = tarjeta = activo.mi_tarjeta_control()
                    form = DetalleTarjetaControlForm(
                        initial={'codigobara': activo.codigogobierno if activo.codigogobierno else activo.codigointerno,
                                 'catalogo': activo.catalogo,
                                 'descripcion': activo.descripcion,
                                 'fechacompra': activo.fechaingreso})
                    form.agregar()
                    data['form'] = form
                    return render(request, "af_activofijo/addtarjeta.html", data)
                except Exception as ex:
                    pass

            elif action == 'edittarjeta':
                try:
                    puede_realizar_accion(request, 'sagest.puede_modificar_tarjeta')
                    data['title'] = u'Editar Envío a Mantenimiento.'
                    data['detalle'] = detalle = DetalleMantenimiento.objects.get(pk=request.GET['id'])
                    activo = detalle.tarjeta.activo
                    form = DetalleTarjetaControlForm(
                        initial={'codigobara': activo.codigogobierno if activo.codigogobierno else activo.codigointerno,
                                 'catalogo': activo.catalogo,
                                 'descripcion': activo.descripcion,
                                 'fechacompra': activo.fechaingreso,
                                 'fechaentrega': detalle.fechaentrega if detalle.fechaentrega else datetime.now().date(),
                                 'mantenimientorealizar': detalle.mantenimientorealizar,
                                 'observacion': detalle.observacion,
                                 'taller': detalle.taller})
                    form.agregar()
                    data['form'] = form
                    return render(request, "af_activofijo/edittarjeta.html", data)
                except Exception as ex:
                    pass

            elif action == 'ingresodemantenimiento':
                try:
                    puede_realizar_accion(request, 'sagest.puede_modificar_tarjeta')
                    data['title'] = u'Ingreso de Mantenimiento y Reparación.'
                    data['detalle'] = detalle = DetalleMantenimiento.objects.get(pk=request.GET['id'])
                    activo = detalle.tarjeta.activo
                    traslaso = None
                    if detalle.detalletrasladomantenimiento:
                        traslado = detalle.detalletrasladomantenimiento.codigotraslado
                    form = DetalleTarjetaControlForm(initial={'mantenimientorealizar': detalle.mantenimientorealizar,
                                                              'mantenimientorealizado': detalle.mantenimientorealizado,
                                                              'codigobara': activo.codigogobierno if activo.codigogobierno else activo.codigointerno,
                                                              'catalogo': activo.catalogo,
                                                              'descripcion': activo.descripcion,
                                                              'fechacompra': activo.fechaingreso,
                                                              'observacion': detalle.observacion,
                                                              'aplicagarantia': detalle.aplicagarantia,
                                                              'taller': detalle.taller,
                                                              'manodeobra': detalle.manodeobra,
                                                              'costomanodeobra': detalle.costomanodeobra,
                                                              'facturamanodeobra': detalle.facturamanodeobra,
                                                              'fechaentrega': detalle.fechaentrega})
                    form.ingreso()
                    data['form'] = form
                    return render(request, "af_activofijo/ingresomantenimiento.html", data)
                except Exception as ex:
                    pass

            elif action == 'histtraslados':
                try:
                    data['title'] = u'Traslados de activos a mantenimiento'
                    search = None
                    ids = None
                    url_vars = f"&action=histtraslados"
                    traslado = TrasladoMantenimiento.objects.filter(status=True).order_by('estado', '-numero')
                    if 's' in request.GET:
                        search = request.GET['s']
                        ss = search.split(' ')
                        if len(ss) == 1:
                            traslado = traslado.filter(Q(numero__icontains=search) |
                                                       Q(detalletrasladomantenimiento__activo__codigointerno=search) |
                                                       Q(asistentelogistica__nombres__icontains=search) |
                                                       Q(asistentelogistica__apellido1__icontains=search) |
                                                       Q(asistentelogistica__apellido2__icontains=search) |
                                                       Q(usuariobienes__nombres__icontains=search) |
                                                       Q(usuariobienes__apellido1__icontains=search) |
                                                       Q(usuariobienes__apellido2__icontains=search) |
                                                       Q(departamentosolicita__nombre__icontains=search)).distinct().order_by(
                                'estado', '-numero')
                        else:
                            traslado = traslado.filter(Q(usuariobienes__apellido1__icontains=ss[0]) &
                                                       Q(usuariobienes__apellido2=ss[1]) |
                                                       Q(asistentelogistica__apellido1__icontains=ss[0]) &
                                                       Q(asistentelogistica__apellido2=ss[1])).distinct().order_by(
                                'estado', '-numero')
                        url_vars += f"&s={search}"
                    elif 'id' in request.GET:
                        ids = request.GET['id']
                        traslado = traslado.filter(id=ids).order_by('estado', '-numero')
                        url_vars += f"&id={ids}"
                    paging = MiPaginador(traslado, 25)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['url_vars'] = url_vars
                    data['traslados'] = page.object_list
                    data['reporte_0'] = obtener_reporte('traslado_activo')
                    data['usuario'] = request.user
                    return render(request, "af_activofijo/traslados.html", data)
                except Exception as ex:
                    pass

            elif action == 'reportebajas':
                try:
                    data['form'] = DescargarBajasAniosForm()
                    data['switchery']=True
                    data['id']=request.GET['id']
                    template = get_template("af_activofijo/modal/formreportebajas.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'movimientos':
                try:
                    data['title'] = u'Movimientos de Activos'
                    searchse = None
                    searchsr = None
                    search = None
                    ids = None
                    url_vars = f"&action=movimientos"
                    traspasos = TraspasoActivo.objects.filter(tipo=2, status=True).order_by('estado', '-numero')
                    if 'se' in request.GET or 'sr' in request.GET or 's' in request.GET:
                        if 'se' in request.GET:
                            searchse = request.GET['se'].strip()
                            ss = searchse.split(' ')
                            if len(ss) == 1:
                                traspasos = traspasos.filter(Q(usuariobienentrega__nombres__icontains=searchse) |
                                                             Q(usuariobienentrega__apellido1__icontains=searchse) |
                                                             Q(usuariobienentrega__apellido2__icontains=searchse),
                                                             detalletraspasoactivo__seleccionado=True).distinct().order_by(
                                    'estado', '-numero')
                            else:
                                traspasos = traspasos.filter(Q(usuariobienentrega__apellido1__icontains=ss[0]) & Q(
                                    usuariobienentrega__apellido2__icontains=ss[1]),
                                                             detalletraspasoactivo__seleccionado=True).distinct().order_by(
                                    'estado', '-numero')
                            url_vars += f"&se={searchse}"
                        if 's' in request.GET:
                            search = request.GET['s']
                            traspasos = traspasos.filter(Q(numero__icontains=search) |
                                                         Q(usuario_creacion__username=search) |
                                                         Q(detalletraspasoactivo__activo__codigointerno=search,
                                                           detalletraspasoactivo__seleccionado=True) |
                                                         Q(detalletraspasoactivo__activo__codigogobierno=search)|
                                                         Q(usuariobienentrega__nombres__icontains=search) |
                                                         Q(usuariobienentrega__apellido1__icontains=search) |
                                                         Q(usuariobienentrega__apellido2__icontains=search) |
                                                         Q(usuariobienrecibe__nombres__icontains=search) |
                                                         Q(usuariobienrecibe__apellido1__icontains=search) |
                                                         Q(usuariobienrecibe__apellido2__icontains=search)).distinct().order_by(
                                'estado', '-numero')
                            url_vars += f"&s={search}"
                        if 'sr' in request.GET:
                            searchsr = request.GET['sr'].strip()
                            ss = searchsr.split(' ')
                            if len(ss) == 1:
                                traspasos = traspasos.filter(Q(usuariobienrecibe__nombres__icontains=searchsr) |
                                                             Q(usuariobienrecibe__apellido1__icontains=searchsr) |
                                                             Q(usuariobienrecibe__apellido2__icontains=searchsr),
                                                             detalletraspasoactivo__seleccionado=True).distinct().order_by(
                                    'estado', '-numero')
                            else:
                                traspasos = traspasos.filter(Q(usuariobienentrega__apellido1__icontains=ss[0]) & Q(
                                    usuariobienentrega__apellido2__icontains=ss[1]),
                                                             detalletraspasoactivo__seleccionado=True).distinct().order_by(
                                    'estado', '-numero')
                            url_vars += f"&sr={searchsr}"
                    elif 'id' in request.GET:
                        ids = int(request.GET['id'])
                        traspasos = traspasos.filter(id=ids).order_by('estado', '-numero')
                        url_vars += f"&id={ids}"
                    paging = MiPaginador(traspasos, 25)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['searchse'] = searchse if searchse else ""
                    data['searchsr'] = searchsr if searchsr else ""
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['traspasos'] = page.object_list
                    data['reporte_0'] = obtener_reporte('traspaso_activo')
                    data['reporte_1'] = obtener_reporte('traspaso_activo_solocustodio')
                    data['usuario'] = request.user
                    data['idpersonamisactivos'] = persona.id
                    request.session['viewactivo'] = 4
                    data['url_vars'] = url_vars
                    return render(request, "af_activofijo/movimientos.html", data)
                except Exception as ex:
                    pass

            elif action == 'actasentrega':
                try:
                    data['title'] = u'Actas de entrega y recepción de Activos'
                    searchse = None
                    searchsr = None
                    url_vars = f"&action=actasentrega"
                    ids = None
                    if 'se' in request.GET or 'sr' in request.GET:
                        traspasos = TraspasoActivo.objects.filter(tipo=1)
                        if 'se' in request.GET:
                            searchse = request.GET['se'].strip()
                            ss = searchse.split(' ')
                            if len(ss) == 1:
                                traspasos = traspasos.filter(
                                    Q(usuariobienentrega__nombres__icontains=searchse, tipo=1) |
                                    Q(usuariobienentrega__apellido1__icontains=searchse, tipo=1) |
                                    Q(usuariobienentrega__apellido2__icontains=searchse, tipo=1)).order_by('-numero')
                            else:
                                traspasos = traspasos.filter(
                                    Q(usuariobienentrega__apellido1__icontains=ss[0], tipo=1) & Q(
                                        usuariobienentrega__apellido2__icontains=ss[1], tipo=1)).order_by('-numero')
                            url_vars += f"&se={searchse}"
                        if 'sr' in request.GET:
                            searchsr = request.GET['sr'].strip()
                            ss = searchsr.split(' ')
                            if len(ss) == 1:
                                traspasos = traspasos.filter(Q(usuariobienrecibe__nombres__icontains=searchsr, tipo=1) |
                                                             Q(usuariobienrecibe__apellido1__icontains=searchsr,
                                                               tipo=1) |
                                                             Q(usuariobienrecibe__apellido2__icontains=searchsr,
                                                               tipo=1)).order_by('-numero')
                            else:
                                traspasos = traspasos.filter(
                                    Q(usuariobienentrega__apellido1__icontains=ss[0], tipo=1) & Q(
                                        usuariobienentrega__apellido2__icontains=ss[1], tipo=1)).order_by('-numero')
                            url_vars += f"&sr={searchsr}"
                    elif 'id' in request.GET:
                        ids = request.GET['id']
                        traspasos = TraspasoActivo.objects.filter(id=ids, tipo=1).order_by('-numero')
                        url_vars += f"&id={ids}"
                    else:
                        traspasos = TraspasoActivo.objects.filter(tipo=1).order_by('-numero')
                    paging = MiPaginador(traspasos, 25)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['searchse'] = searchse if searchse else ""
                    data['searchsr'] = searchsr if searchsr else ""
                    data['ids'] = ids if ids else ""
                    data['url_vars'] = url_vars
                    data['traspasos'] = page.object_list
                    data['reporte_0'] = obtener_reporte('acta_entrega_libro')
                    data['reporte_1'] = obtener_reporte('acta_entrega_vehiculo')
                    data['reporte_2'] = obtener_reporte('acta_entrega_otro')
                    return render(request, "af_activofijo/actasentrega.html", data)
                except Exception as ex:
                    pass

            elif action == 'tarjeta':
                try:
                    data['title'] = u'Detalle Tarjeta de control'
                    search = None
                    ids = None
                    idt = None
                    url_vars = f"&action=tarjeta"
                    if 's' in request.GET:
                        search = request.GET['s']
                        detalle = TarjetaControl.objects.filter(Q(activo__descripcion__icontains=search) |
                                                                Q(activo__codigointerno=search) |
                                                                Q(activo__codigogobierno=search) |
                                                                Q(numero=search), status=True).distinct().order_by(
                            '-numero')
                        url_vars += f"&s={search}"
                    elif 'id' in request.GET:
                        ids = request.GET['id']
                        detalle = TarjetaControl.objects.filter(id=ids, status=True).order_by('-numero')
                        url_vars += f"&id={ids}"
                    else:
                        detalle = TarjetaControl.objects.filter(status=True).order_by('-numero')
                    paging = MiPaginador(detalle, 25)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['idt'] = idt if idt else ""
                    data['url_vars'] = url_vars
                    data['detalles'] = page.object_list
                    data['reporte_0'] = obtener_reporte('tarjeta_control')
                    return render(request, "af_activofijo/tarjeta.html", data)
                except Exception as ex:
                    pass

            elif action == 'detalletarjeta':
                try:
                    data['title'] = u'Detalle Tarjeta de control'
                    data['tarjeta'] = tarjeta = TarjetaControl.objects.get(pk=int(request.GET['id']))
                    data['detalles'] = tarjeta.detallemantenimiento_set.all()
                    return render(request, "af_activofijo/detalletarjeta.html", data)
                except Exception as ex:
                    pass

            elif action == 'confirmacion':
                try:
                    data['title'] = u'Confirmar traslado'
                    data['traslado'] = TrasladoMantenimiento.objects.get(pk=int(request.GET['id']))
                    return render(request, "af_activofijo/confirmacion.html", data)
                except:
                    pass

            elif action == 'excelmangarantia':
                try:
                    __author__ = 'Unemi'
                    style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
                    style_nb = easyxf('font: name Times New Roman, color-index blue, bold on',
                                      num_format_str='#,##0.00')
                    style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
                    title = easyxf(
                        'font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                    style1 = easyxf(num_format_str='D-MMM-YY')
                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('exp_xls_post_part')
                    ws.write_merge(0, 0, 0, 10, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                    response = HttpResponse(content_type="application/ms-excel")
                    fech_ini = request.GET['fechainicio']
                    fech_fin = request.GET['fechafin']
                    tip_act = request.GET['tact']
                    response[
                        'Content-Disposition'] = 'attachment; filename=listadomangarantia' + random.randint(1,
                                                                                                            10000).__str__() + '.xls'

                    columns = [
                        (u"ACTIVO", 15000),
                        (u"PROVEEDOR", 15000),
                        (u"FECHA DE EJECUCION", 3000),
                        (u"TIPO DE ACTIVO", 10000),
                        (u"VALOR", 3000),
                        (u"Hora(s)", 3000),
                        (u"Minuto(s)", 3000),
                        (u"FUNCIONA AL RECIBIR", 8000),
                        (u"FUNCIONA AL ENTREGAR", 8000),
                        (u"USUARIO ENTREGO EQUIPO", 8000),
                        (u"OBSERVACION", 10000),
                        (u"TAREAS DE LIMPIEZA", 10000),
                        (u"DAÑOS ENCONTRADOS", 10000),
                        (u"USUARIO RESP.", 8000)
                    ]
                    row_num = 3
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], font_style)
                        ws.col(col_num).width = columns[col_num][1]
                    date_format = xlwt.XFStyle()
                    date_format.num_format_str = 'yyyy/mm/dd'
                    listamangarantia = MantenimientosActivosGarantia.objects.db_manager('sga_select').filter(
                        status=True, fechainicio__range=[fech_ini, fech_fin], tipoactivo__id=tip_act)
                    row_num = 4
                    for lista in listamangarantia:
                        i = 0
                        campo1 = lista.activofijo.descripcion + ' ' + lista.activofijo.codigointerno + ' ' + lista.activofijo.codigogobierno
                        if lista.proveedor:
                            campo2 = lista.proveedor.nombre
                        else:
                            campo2 = ''
                        campo3 = lista.fechainicio
                        campo5 = lista.valor
                        campo6 = lista.personacreador()
                        campo7 = lista.horamax
                        campo8 = lista.minutomax
                        if lista.estfrec:
                            campo9 = 'SI'
                        else:
                            campo9 = 'NO'
                        if lista.estfent:
                            campo10 = 'SI'
                        else:
                            campo10 = 'NO'
                        if lista.estusu:
                            campo14 = 'SI'
                        else:
                            campo14 = 'NO'
                        campo11 = lista.observacion
                        listatareas = []
                        danioenc = []
                        if lista.tareasactivospreventivosgarantialimp_set.values_list('grupos__descripcion',
                                                                                      flat=True).filter(
                            status=True).exists():
                            campo12 = lista.tareasactivospreventivosgarantialimp_set.values_list('grupos__descripcion',
                                                                                                 flat=True).filter(
                                status=True)
                            for listadocampo12 in campo12:
                                listatareas.append(listadocampo12)
                        else:
                            listatareas.append('NINGUNA')
                        if lista.tareasactivospreventivosgarantiaerr_set.values_list('grupos__descripcion',
                                                                                     flat=True).filter(
                            status=True).exists():
                            campo13 = lista.tareasactivospreventivosgarantiaerr_set.values_list('grupos__descripcion',
                                                                                                flat=True).filter(
                                status=True)
                            for listadocampo13 in campo13:
                                danioenc.append(listadocampo13)
                        else:
                            danioenc.append('NINGUNA')
                        campo15 = lista.tipoactivo.descripcion
                        ws.write(row_num, 0, campo1, font_style2)
                        ws.write(row_num, 1, campo2, font_style2)
                        ws.write(row_num, 2, campo3, date_format)
                        ws.write(row_num, 3, campo15, font_style2)
                        ws.write(row_num, 4, campo5, font_style2)
                        ws.write(row_num, 5, campo7, font_style2)
                        ws.write(row_num, 6, campo8, font_style2)
                        ws.write(row_num, 7, campo9, font_style2)
                        ws.write(row_num, 8, campo10, font_style2)
                        ws.write(row_num, 9, campo14, font_style2)
                        ws.write(row_num, 10, campo11, font_style2)
                        ws.write(row_num, 11, str(listatareas), font_style2)
                        ws.write(row_num, 12, str(danioenc), font_style2)
                        ws.write(row_num, 13, campo6, font_style2)
                        row_num += 1
                    wb.save(response)
                    return response
                except Exception as ex:
                    pass

            elif action == 'excelmanpreventivos':
                try:
                    __author__ = 'Unemi'
                    style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
                    style_nb = easyxf('font: name Times New Roman, color-index blue, bold on',
                                      num_format_str='#,##0.00')
                    style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
                    title = easyxf(
                        'font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                    style1 = easyxf(num_format_str='D-MMM-YY')
                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('exp_xls_post_part')
                    # ws.write_merge(0, 0, 0, 10, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=listadomanpreventivos' + random.randint(1,
                                                                                                                    10000).__str__() + '.xls'
                    fech_ini = request.GET['fechainicio']
                    fech_fin = request.GET['fechafin']
                    tip_act = request.GET['tact']
                    columns = [
                        (u"Activo", 10000),
                        (u"Codigo Interno", 4000),
                        (u"Codigo Gobierno", 4000),
                        (u"Monitor", 4000),
                        (u"Teclado", 4000),
                        (u"Mouse", 4000),
                        (u"Mantenimiento", 4000),
                        (u"Procesador", 4000),
                        (u"Memoria", 4000),
                        (u"Disco duro", 4000),
                        (u"Particiones", 4000),
                        (u"Sistema operativo", 4000),
                        (u"Arquitectura", 4000),
                        (u"Service pack", 4000),
                        (u"Fecha mantenimiento", 4000),
                        (u"Funciona recibe", 4000),
                        (u"Funciona entrega", 4000),
                        (u"Tareas realizadas", 15000),
                        (u"Tareas no realizadas", 15000),
                        (u"Observaciones", 10000),
                        (u"Usuario Ejecuto Mant.", 10000)
                    ]

                    row_num = 0
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], font_style)
                        ws.col(col_num).width = columns[col_num][1]
                    date_format = xlwt.XFStyle()
                    date_format.num_format_str = 'yyyy-mm-dd'
                    date_formatreverse = xlwt.XFStyle()
                    date_formatreverse.num_format_str = 'dd/mm/yyyy'
                    listadomanpreventivos = MantenimientosActivosPreventivos.objects.select_related().filter(
                        fecha__gte=fech_ini, fecha__lte=fech_fin, tipoactivo__id=tip_act, status=True)
                    row_num = 1
                    i = 0
                    for listado in listadomanpreventivos:
                        i += 1
                        campo1 = listado.activofijo.descripcion + ' ' + listado.activofijo.codigointerno + ' ' + listado.activofijo.codigogobierno
                        campo2 = listado.activofijo.codigointerno
                        campo3 = listado.activofijo.codigogobierno
                        campo4 = listado.monitor
                        campo5 = listado.teclado
                        campo6 = listado.mouse
                        campo7 = listado.get_tipomantenimiento_display()
                        campo8 = listado.procesador
                        campo9 = listado.memoria
                        campo10 = listado.discoduro
                        campo11 = listado.particiones
                        campo12 = listado.sistemaoperativo
                        campo13 = listado.arquitectura
                        campo14 = listado.service
                        campo15 = listado.fecha
                        campo16 = 'NO'
                        campo17 = 'NO'
                        if listado.funcionarecibe:
                            campo16 = 'SI'
                        if listado.funcionaentrega:
                            campo17 = 'SI'
                        listatareas = []
                        listatareasno = []
                        if listado.tareasactivospreventivos_set.values_list('grupos__descripcion', flat=True).filter(
                                status=True).exists():
                            campo18 = listado.tareasactivospreventivos_set.values_list('grupos__descripcion',
                                                                                       flat=True).filter(status=True)
                            for listadocampo18 in campo18:
                                listatareas.append(listadocampo18)
                        else:
                            listatareas = ''
                        campo19 = MantenimientoGruCategoria.objects.values_list('descripcion', flat=True).filter(
                            grupocategoria=listado.tipoactivo, status=True).exclude(
                            pk__in=listado.tareasactivospreventivos_set.values_list('grupos__id', flat=True).filter(
                                status=True))
                        for listadocampo19 in campo19:
                            listatareasno.append(listadocampo19)
                        campo20 = listado.observaciones
                        campo21 = listado.personacreador()
                        ws.write(row_num, 0, campo1, font_style2)
                        ws.write(row_num, 1, campo2, font_style2)
                        ws.write(row_num, 2, campo3, font_style2)
                        ws.write(row_num, 3, campo4, font_style2)
                        ws.write(row_num, 4, campo5, font_style2)
                        ws.write(row_num, 5, campo6, font_style2)
                        ws.write(row_num, 6, campo7, font_style2)
                        ws.write(row_num, 7, campo8, font_style2)
                        ws.write(row_num, 8, campo9, font_style2)
                        ws.write(row_num, 9, campo10, font_style2)
                        ws.write(row_num, 10, campo11, font_style2)
                        ws.write(row_num, 11, campo12, font_style2)
                        ws.write(row_num, 12, campo13, font_style2)
                        ws.write(row_num, 13, campo14, font_style2)
                        ws.write(row_num, 14, campo15, date_format)
                        ws.write(row_num, 15, campo16, font_style2)
                        ws.write(row_num, 16, campo17, font_style2)
                        ws.write(row_num, 17, str(listatareas), font_style2)
                        ws.write(row_num, 18, str(listatareasno), font_style2)
                        ws.write(row_num, 19, campo20, font_style2)
                        ws.write(row_num, 20, campo21, font_style2)
                        row_num += 1
                    wb.save(response)
                    return response
                except Exception as ex:
                    pass

            elif action == 'atenciontecnologico':
                try:
                    data['title'] = u'Equipo Tecnológico'
                    search = None
                    ids = None
                    data['codigo'] = codigo = int(request.GET['codigo']) if 'codigo' in request.GET else 0
                    data['baja'] = baja = int(request.GET['baja']) if 'baja' in request.GET else 0
                    data['rangosemaforo'] = RangoVidaUtil.objects.filter(status=True).order_by('anio', 'descripcion')
                    data['grupocatalogo'] = GruposCategoria.objects.filter(status=True)
                    if codigo == 0:
                        activos = ActivoFijo.objects.filter(
                            Q(archivobaja__isnull=True) | Q(archivobaja=''), catalogo__equipoelectronico=True,
                            catalogo__status=True, status=True).order_by('descripcion')
                    else:
                        activos = ActivoFijo.objects.filter(
                            Q(archivobaja__isnull=True) | Q(archivobaja=''), catalogo__equipoelectronico=True,
                            catalogo__status=True, catalogo__grupo__id=codigo, status=True).order_by(
                            'descripcion')
                    if 's' in request.GET:
                        search = request.GET['s'].strip()
                        ss = search.split(' ')
                        if len(ss) == 1:
                            activos = activos.filter(Q(id=search) | Q(codigogobierno=search) | Q(codigointerno=search),
                                                     status=True)
                    if baja == 1 or baja == 0:
                        activos = activos.filter(statusactivo=1)
                    else:
                        if baja == 2:
                            activos = activos.filter(statusactivo=2)
                    data['totales'] = activos.values('id').count()

                    paging = MiPaginador(activos, 15)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['listadocatalogo'] = page.object_list
                    return render(request, "af_activofijo/histatenciontecnologico.html", data)
                except Exception as ex:
                    pass

            elif action == 'configurardirector':
                try:
                    data['title'] = u'Configurar director en archivo de baja'
                    search = None
                    ids = None
                    directores = DirectorResponsableBaja.objects.filter(status=True)
                    if 's' in request.GET:
                        search = request.GET['s'].strip()
                        ss = search.split(' ')
                        if len(ss) == 1:
                            directores = directores.filter(Q(id=search) | Q(responsable__nombres__icontains=search) |
                                                           Q(responsable__apellido1__icontains=search) |
                                                           Q(responsable__apellido2__icontains=search) |
                                                           Q(responsable__cedula__icontains=search)
                                                           , status=True)
                    paging = MiPaginador(directores, 15)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['directores'] = page.object_list
                    return render(request, "af_activofijo/viewconfiguradirector.html", data)
                except Exception as ex:
                    pass

            if action == 'adddirector':
                try:
                    data['title'] = u'Adicionar director'
                    data['form2'] = DirectorResponsableBajaForm()
                    template = get_template("af_activofijo/modal/formadicionardirector.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            if action == 'editdirector':
                try:
                    data['id'] = request.GET['id']
                    data['tipo'] = director = DirectorResponsableBaja.objects.get(pk=request.GET['id'])
                    data['form2'] = DirectorResponsableBaja2Form(initial={'actual': director.actual,
                                                                          'fechainicio': str(director.fechainicio),
                                                                          'fechafin': str(director.fechafin)
                                                                          })
                    template = get_template("af_activofijo/modal/formadicionardirector.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'attecpdf':
                try:
                    data = {}
                    data['fechaactual'] = datetime.now()
                    data['activofijo'] = activofijo = ActivoFijo.objects.get(pk=int(request.GET['id']))
                    data['detallemantenimiento'] = HdDetalle_Incidente.objects.filter(incidente__activo=activofijo,
                                                                                      status=True,
                                                                                      incidente__status=True,
                                                                                      resolucion__isnull=False).exclude(
                        resolucion__exact='')
                    data['mantenimientopreventivo'] = activofijo.mantenimientosactivospreventivos_set.filter(
                        status=True)
                    data[
                        'mantenimientogarantia'] = mantenimientogarantia = activofijo.mantenimientosactivosgarantia_set.filter(
                        status=True)
                    data['costomantenimientogarantia'] = \
                        mantenimientogarantia.filter(status=True).aggregate(cantidad=Sum('valor'))['cantidad']
                    return conviert_html_to_pdf('af_activofijo/histtecnologico_pdf.html',
                                                {'pagesize': 'A4', 'data': data})
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Problemas al ejecutar el reporte. %s" % ex})

            elif action == 'histsingarpdf':
                try:
                    data = {}
                    data['fechaactual'] = datetime.now()
                    data['activofijo'] = activofijo = ActivoFijo.objects.get(pk=int(request.GET['id']))
                    data['mantenimiento'] = mantenimiento = activofijo.mantenimientosactivospreventivos_set.get(
                        status=True, id=int(request.GET['idh']))
                    data['tareasmantenimiento'] = MantenimientoGruCategoria.objects.filter(
                        grupocategoria=mantenimiento.tipoactivo, status=True)
                    data['tareasactivo'] = TareasActivosPreventivos.objects.values_list('grupos_id', flat=True).filter(
                        mantenimiento=mantenimiento, status=True)
                    data['ppmantenimiento'] = HdPiezaPartes.objects.filter(grupocategoria=mantenimiento.tipoactivo,
                                                                           status=True)
                    data['piezaparteactivo'] = PiezaParteActivosPreventivos.objects.filter(mantenimiento=mantenimiento,
                                                                                           status=True)
                    data['daniomantenimiento'] = MantenimientoGruDanios.objects.filter(
                        grupocategoria=mantenimiento.tipoactivo, status=True)
                    data['danioactivo'] = TareasActivosPreventivosDanios.objects.filter(mantenimiento=mantenimiento,
                                                                                        status=True)
                    if mantenimiento.nuevo:
                        return conviert_html_to_pdf('af_activofijo/histmantsingar_pdfv2.html',
                                                    {'pagesize': 'A4', 'data': data})
                    else:
                        return conviert_html_to_pdf('af_activofijo/histmantsingar_pdf.html',
                                                    {'pagesize': 'A4', 'data': data})
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Problemas al ejecutar el reporte. %s" % ex})

            elif action == 'histincidentepdf':
                try:
                    data = {}
                    data['mantenimiento'] = mantenimiento = HdDetalle_Incidente.objects.get(id=request.GET['idh'])
                    return conviert_html_to_pdf('af_activofijo/histincidentes_pdf.html',
                                                {'pagesize': 'A4', 'data': data})
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Problemas al ejecutar el reporte. %s" % ex})

            elif action == 'histcongarpdf':
                try:
                    data = {}
                    data['fechaactual'] = datetime.now()
                    data['activofijo'] = activofijo = ActivoFijo.objects.get(pk=int(request.GET['id']))
                    data['mantenimiento'] = mantenimiento = activofijo.mantenimientosactivosgarantia_set.get(
                        status=True, id=int(request.GET['idh']))
                    data['tareasmantenimiento'] = MantenimientoGruCategoriaGarantiaLimp.objects.filter(
                        grupocategoria=mantenimiento.tipoactivo, status=True)
                    data['tareasactivo'] = TareasActivosPreventivosGarantiaLimp.objects.values_list('grupos_id',
                                                                                                    flat=True).filter(
                        mantenimiento=mantenimiento, status=True)
                    data['danioenc'] = MantenimientoGruCategoriaGarantiaErr.objects.filter(
                        grupocategoria=mantenimiento.tipoactivo, status=True)
                    data['danios'] = TareasActivosPreventivosGarantiaErr.objects.values_list('grupos_id',
                                                                                             flat=True).filter(
                        mantenimiento=mantenimiento, status=True)
                    return conviert_html_to_pdf('af_activofijo/histmantcongar_pdf.html',
                                                {'pagesize': 'A4', 'data': data})
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Problemas al ejecutar el reporte. %s" % ex})

            elif action == 'addgrupo':
                try:
                    data['form2'] = GrupoBienForm()
                    data['id'] = request.GET['id']
                    data['action'] = request.GET['action']
                    template = get_template("af_activofijo/modal/formadicionar.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Error mostrar el formulario. %s" % ex})

            elif action == 'reportebaja':
                try:
                    activos = ActivoFijo.objects.filter(status=True, procesobaja=True, statusactivo=1)
                    __author__ = 'Unemi'

                    output = io.BytesIO()
                    workbook = xlsxwriter.Workbook(output)
                    ws = workbook.add_worksheet('Listado')
                    ws.set_column(0, 2, 20)
                    ws.set_column(3, 9, 60)
                    ws.set_column(10, 12, 20)
                    ws.set_column(13, 13, 10)  # Columna En uso
                    ws.set_column(14, 14, 20)  # Columna Costo
                    ws.set_column(15, 15, 60)  # Columna Reportado por
                    ws.set_column(16, 16, 20)  # Columna Fecha que se reporto
                    ws.set_column(17, 17, 30)

                    formatoceldagris = workbook.add_format(
                        {'align': 'center', 'border': 1, 'text_wrap': True, 'fg_color': '#B6BFC0'})

                    ws.write(0, 0, 'Nro.', formatoceldagris)
                    ws.write(0, 1, 'Cod. gobierno', formatoceldagris)
                    ws.write(0, 2, 'Cod. interno', formatoceldagris)
                    ws.write(0, 3, 'Bien', formatoceldagris)
                    ws.write(0, 4, 'Descripción', formatoceldagris)
                    ws.write(0, 5, 'Usuario', formatoceldagris)
                    ws.write(0, 6, 'Ubicación', formatoceldagris)
                    ws.write(0, 7, 'Cuenta contable', formatoceldagris)
                    ws.write(0, 8, 'Marca', formatoceldagris)
                    ws.write(0, 9, 'Modelo', formatoceldagris)
                    ws.write(0, 10, 'Serie', formatoceldagris)
                    ws.write(0, 11, 'Estado', formatoceldagris)
                    ws.write(0, 12, 'Condición', formatoceldagris)
                    ws.write(0, 13, '¿En uso?', formatoceldagris)
                    ws.write(0, 14, 'Costo', formatoceldagris)
                    ws.write(0, 15, 'Reportado por', formatoceldagris)
                    ws.write(0, 16, 'Fecha que se reporto', formatoceldagris)
                    ws.write(0, 17, 'Ubicación en bodega', formatoceldagris)
                    cont = 1
                    for activo in activos:
                        historial = HistorialEstadoActivo.objects.filter(activo=activo, status=True, estado_id=3).order_by('-id').first()
                        reportado_por = 'S/R'
                        fecha = 'S/F'
                        condicionestado = activo.get_condicionestado_display() if activo.condicionestado else ''
                        enuso = 'SI' if activo.enuso else ('NO' if activo.enuso is not None else '')
                        if historial:
                            reportado_por = Persona.objects.get(usuario=historial.usuario_creacion).nombre_completo_minus()
                            fecha = historial.fecha_creacion
                        ws.write(cont, 0, str(cont))
                        ws.write(cont, 1, str(activo.codigogobierno))
                        ws.write(cont, 2, str(activo.codigointerno))
                        ws.write(cont, 3, str(activo.catalogo))
                        ws.write(cont, 4, str(activo.descripcion))
                        ws.write(cont, 5, str(activo.responsable))
                        ws.write(cont, 6, str(activo.ubicacion))
                        ws.write(cont, 7, str(activo.cuentacontable))
                        ws.write(cont, 8, str(activo.marca))
                        ws.write(cont, 9, str(activo.modelo))
                        ws.write(cont, 10, str(activo.serie))
                        ws.write(cont, 11, str(activo.estado))
                        ws.write(cont, 12, condicionestado)
                        ws.write(cont, 13, enuso)
                        ws.write(cont, 14, str(activo.costo))
                        ws.write(cont, 15, str(reportado_por))
                        ws.write(cont, 16, str(fecha))
                        ws.write(cont, 17, str(activo.ubicacionbodega))
                        cont += 1
                    workbook.close()
                    output.seek(0)
                    filename = 'activos_proceso_baja.xlsx'
                    response = HttpResponse(output,
                                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    response['Content-Disposition'] = 'attachment; filename=%s' % filename
                    return response
                except Exception as ex:
                    pass

            elif action == 'reportemantenimientoactivos':
                try:
                    trasladomantenimientoactivos = TrasladoMantenimiento.objects.filter(status=True).order_by('estado',
                                                                                                              '-numero')
                    data['listadomantenimientoactivos'] = trasladomantenimientoactivos
                    data['totalmantenimientoactivos'] = trasladomantenimientoactivos.count()
                    data['fecha'] = datetime.now().date()
                    return conviert_html_to_pdf('af_activofijo/reportemantenimientoactivos.html',
                                                {'pagesize': 'A4', 'data': data})
                except Exception as ex:
                    pass

            elif action == 'firmaractatraspaso':
                try:
                    solicitudtraspaso = SolicitudTraspasoActivos.objects.get(id=int(request.GET['id']))
                    traspasoactivoporfirmar = TraspasoActivo.objects.get(id=solicitudtraspaso.traspasoactivofijo_id)
                    data['action_firma'] = 'firmardocumento'
                    data['id_objeto'] = traspasoactivoporfirmar.id
                    data['archivo'] = archivoporfirmar = '/media/' + traspasoactivoporfirmar.traspasoactivofirma.name
                    data['url_archivo'] = '{}{}'.format(dominio_sistema,
                                                        '/media/' + traspasoactivoporfirmar.traspasoactivofirma.name)
                    template = get_template("formfirmaelectronica.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            #CONSTATACION DE ACTIVOS FIJOS
            elif action == 'periodoconstatacion':
                try:
                    data['title'] = u'Periodos de constatación'
                    data['subtitle'] = u'Listado de periodos de constatación de activos fijos'
                    search, url_vars, filtro = request.GET.get('s', ''), f'&action={action}', Q(status=True)
                    if search:
                        data['s'] = search
                        filtro = filtro & Q(nombre__icontains=search)
                        url_vars += f'&s={search}'
                    periodo = PeriodoConstatacionAF.objects.filter(filtro).order_by('-id')
                    paging = MiPaginador(periodo, 10)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['url_vars'] = url_vars
                    data['listado'] = page.object_list
                    request.session['viewactivo'] = 2
                    return render(request, "af_activofijo/periodosconstatacion.html", data)
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'addperiodo':
                try:
                    form = PeriodoConstatacionAFForm()
                    data['form'] = form
                    data['switchery']=True
                    template = get_template('ajaxformmodal.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'editperiodo':
                try:
                    id=int(encrypt(request.GET['id']))
                    periodo=PeriodoConstatacionAF.objects.get(id=id)
                    form = PeriodoConstatacionAFForm(initial=model_to_dict(periodo))
                    data['id']=id
                    data['form'] = form
                    data['switchery'] = True
                    template = get_template('ajaxformmodal.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'activosconstatar':
                try:
                    data['title'] = u'Constatación de activos'
                    data['periodo_c'] = periodo = PeriodoConstatacionAF.objects.get(id=int(encrypt(request.GET['id'])))
                    filtro_excl=[]
                    constatador, estado, constatacion, search, url_vars, filtro =  request.GET.get('constatador', ''), \
                                                                                   request.GET.get('estado', ''), \
                                                                                   request.GET.get('constatacion', ''), \
                                                                                   request.GET.get('s', ''), \
                                                                                   f'&action={action}&id={request.GET["id"]}', \
                                                                                   Q(status=True, statusactivo=1)
                    if periodo.cerrado:
                        filtro &= Q(fecha_creacion__lte=periodo.fechacierre)

                    if estado:
                        data['estado'] = int(estado)
                        url_vars += "&estado={}".format(estado)
                        filtro = filtro & Q(estado_id=estado)

                    if constatacion:
                        data['constatacion'] = constatacion = int(constatacion)
                        url_vars += "&constatacion={}".format(constatacion)
                        ids_a = periodo.detalle_constatacion().values_list('activo_id', flat=True)
                        if constatacion == 1:
                            filtro = filtro & Q(id__in=ids_a)
                        elif constatacion == 2:
                            filtro_excl = ids_a

                    if constatador:
                        data['constatador']=constatador=int(constatador)
                        url_vars += "&constatador={}".format(constatador)
                        ids_a=periodo.detalle_constatacion().filter(responsable_id=constatador).values_list('activo_id',flat=True)
                        filtro = filtro & Q(id__in=ids_a)
                    # if clase:
                    #     data['clase'] = int(clase)
                    #     url_vars += "&clase={}".format(clase)
                    #     filtro = filtro & Q(clasebien=clase)

                    # if cuenta:
                    #     data['cuenta'] = int(cuenta)
                    #     url_vars += "&cuenta={}".format(cuenta)
                    #     filtro = filtro & Q(cuentacontable=cuenta)

                    if search:
                        ss = search.split(' ')
                        data['s'] = search
                        url_vars += "&s={}".format(search)
                        if len(ss) == 1:
                            filtro = filtro & (Q(codigogobierno=search) |
                                               Q(serie=search) |
                                               Q(responsable__cedula=search) |
                                               Q(responsable__pasaporte=search) |
                                               Q(responsable__apellido1__icontains=search) |
                                               Q(responsable__apellido2__icontains=search) |
                                               Q(responsable__nombres__icontains=search) |
                                               Q(codigointerno=search))
                        else:
                            filtro = filtro & (Q(responsable__apellido1__icontains=ss[0]) &
                                               Q(responsable__apellido2__icontains=ss[1]) |
                                               Q(responsable__nombres__icontains=search) |
                                               Q(descripcion__icontains=search))

                    activos = ActivoFijo.objects.select_related().filter(filtro).exclude(id__in=filtro_excl).order_by('-id')
                    paging = MiPaginador(activos, 10)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    # data['clases'] = CLASE_BIEN
                    # data['cuentas'] = CuentaContable.objects.filter(status=True, activosfijos=True)
                    data['estados'] = EstadoProducto.objects.filter(status=True).exclude(id__in=[4,5])
                    data['constatadores']=periodo.constatadores()
                    data['totales'] = periodo.totales()
                    data['url_vars'] = url_vars
                    data['listado'] = page.object_list
                    request.session['viewactivo'] = 2
                    return render(request, "af_activofijo/activosconstatar.html", data)
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'constatar':
                try:
                    id_p, id_a, id_dc=encrypt_id(request.GET['idp']), encrypt_id(request.GET['idex']), encrypt_id(request.GET['id'])
                    activo=ActivoFijo.objects.get(id=id_a)
                    detalle_c=DetalleConstatacionFisica.objects.filter(activo_id=id_a, codigoconstatacion__periodo_id=id_p, status=True).first()
                    if detalle_c:
                        if not detalle_c.responsable.id == persona.id:
                            messages.error(request, f'Activo que intento constatar ya fue constatado por: {detalle_c.responsable.nombre_completo_minus()}')
                            return JsonResponse({"result": False, 'reload': True})
                        form=ConstatacionFisicaForm(initial=model_to_dict(detalle_c))
                        if detalle_c.usuariobienes:
                            form.fields['usuariobienes'].queryset = Persona.objects.filter(id=detalle_c.usuariobienes.id)
                        else:
                            form.fields['usuariobienes'].queryset = Persona.objects.none()
                        data['edit']=True
                    else:
                        form = ConstatacionFisicaForm(initial={'ubicacionbienes': activo.ubicacion,
                                                               'estadoactual': activo.estado,
                                                               'condicionestado': activo.condicionestado,
                                                               'encontrado': True, 'enuso': True})
                        form.fields['usuariobienes'].queryset = Persona.objects.none()
                    data['activo']=activo
                    data['periodo']=id_p
                    data['form'] = form
                    template = get_template('af_activofijo/modal/formconstatar.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'detalle_bajas':
                try:
                    data['baja'] = baja = BajaActivo.objects.get(pk=int(request.GET['id']))
                    data['detalles'] = baja.detallebajaactivo_set.all()
                    template = get_template("af_activofijo/detallebaja.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'buscarpersonas':
                try:
                    resp = filtro_persona_select(request)
                    return HttpResponse(json.dumps({'status': True, 'results': resp}))
                except Exception as ex:
                    pass

            elif action == 'movimiento':
                try:
                    id_a = encrypt_id(request.GET['id'])
                    activo=ActivoFijo.objects.get(id=id_a)
                    data['activo']=activo
                    data["detalles"] = activo.detalletraspasoactivo_set.filter(status=True).order_by('-codigotraspaso__fecha')
                    data['reporte_0'] = obtener_reporte('historial_activos')
                    template = get_template('af_activofijo/modal/movimientos.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'detalleconstatacion':
                try:
                    data['detalle_c'] = DetalleConstatacionFisica.objects.get(pk=encrypt_id(request.GET['id']))
                    template = get_template("af_activofijo/modal/detalleconstatacion.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, "mensaje": u"Error al obtener los datos."})

            elif action == 'reporteconstatacion':
                try:
                    data['idp'] = idp = int(encrypt(request.GET['id']))
                    periodo_c = PeriodoConstatacionAF.objects.get(id=idp)
                    filtro = Q(status=True, codigoconstatacion__periodo_id=idp)
                    if periodo_c.cerrado:
                        filtro &= Q(activo__fecha_creacion__lte=periodo_c.fechacierre)
                    constataciones_base = DetalleConstatacionFisica.objects.filter(filtro)
                    data['title'] = u'Constataciones por responsable'
                    desde = request.GET.get('desde', '')
                    hasta = request.GET.get('hasta', '')
                    after = request.GET.get('after', '')
                    before = request.GET.get('before', '')
                    url_vars = f"&action={action}"
                    if after:
                        filtro = filtro & Q(fecha_creacion__date__gt=after)
                    if before:
                        filtro = filtro & Q(fecha_creacion__date__lt=before)
                        fechas = constataciones_base.filter(filtro).values_list('fecha_creacion__date',flat=True).distinct().order_by(
                            '-fecha_creacion__date')
                    else:
                        fechas = constataciones_base.filter(filtro).values_list('fecha_creacion__date',flat=True).distinct().order_by(
                            'fecha_creacion__date')

                    if desde and hasta:
                        filtro = filtro & (Q(fecha_creacion__range=[desde, hasta]))
                        lenafter = 0
                        if not url_vars == '':
                            url_vars += f"&desde={desde}&hasta={hasta}"
                            data['url_vars'] = url_vars
                    else:
                        lenfecha = fechas.__len__()
                        lenafter = lenfecha
                        lenbefore = lenfecha
                        fec = 7
                        if fechas.__len__() > 0:
                            if lenfecha >= fec:
                                fechas = fechas[:fec]
                                lenfecha = fechas.__len__()
                            if after:
                                lenafter = lenafter - fec
                                lenbefore = lenbefore + fec
                            if before:
                                lenbefore = lenbefore - fec
                                lenafter = lenafter + fec
                                fechas.reverse()
                            if 'desde' not in request.GET and 'hasta' not in request.GET:
                                desde = fechas[0].strftime('%Y-%m-%d')  # Formato YYYY-MM-DD
                                hasta = fechas[lenfecha - 1].strftime('%Y-%m-%d')  # Formato DD/MM/YYYY
                                filtro = filtro & (Q(fecha_creacion__date__range=[desde, hasta]))
                                url_vars = ''
                    if 'after' not in request.GET and 'before' not in request.GET:
                        lenbefore = 0

                    data['lenafter'] = lenafter
                    data['lenbefore'] = lenbefore
                    data['desde'] = desde
                    data['hasta'] = hasta
                    data['fechas'] = fechas = constataciones_base.filter(filtro).values_list(
                        'fecha_creacion__date', flat=True).distinct().order_by('fecha_creacion__date')
                    idresp = constataciones_base.filter(status=True).values_list('responsable_id', flat=True).order_by(
                        'responsable_id')
                    data['respons'] = respons = Persona.objects.filter(id__in=idresp)
                    data['estados'] = EstadoProducto.objects.filter(status=True).exclude(id__in=[4,5])
                    periodo = PeriodoConstatacionAF.objects.get(id=idp)
                    data['constatadores'] = periodo.constatadores()
                    data['totalporfecha'] = []
                    for fecha in fechas:
                        t = constataciones_base.filter(status=True, fecha_creacion__date=fecha,responsable_id__in=respons).count()
                        data['totalporfecha'].append(t)

                    return render(request, 'inventario_activofijo/constatacionesresponsables.html', data)
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'historicoactivoconstatacion':
                try:
                    data['title'] = u'Historial Estados de Activos por Constatación'
                    data['subtitle'] = u'Estados de Activos Fijos según constataciones físicas'
                    search, url_vars, filtro = request.GET.get('s', ''), f'&action={action}', Q(status=True)
                    fechadesde, fechahasta = request.GET.get('fechadesde', ''), request.GET.get('fechahasta', '')
                    fecha = datetime.now().date()

                    if fechadesde:
                        fechadesde = datetime.strptime(fechadesde, '%Y-%m-%d').date()
                        fechahasta = datetime.strptime(fechahasta, '%Y-%m-%d').date()
                    else:
                        fechadesde = fechahasta = fecha

                    url_vars += f'&fechadesde={fechadesde}&fechahasta={fechahasta}'

                    filtro = filtro & Q(codigoconstatacion__estado=2, codigoconstatacion__fechainicio__range=[fechadesde, fechahasta], codigoconstatacion__fechafin__range=[fechadesde, fechahasta])

                    if search:
                        data['s'] = search
                        filtro = filtro & (Q(activo__codigogobierno=search) | Q(activo__codigointerno=search))
                        url_vars += f'&s={search}'

                    detallesconstatacion = DetalleConstatacionFisica.objects.filter(filtro).order_by('id')

                    paging = MiPaginador(detallesconstatacion, 10)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['url_vars'] = url_vars
                    data['detallesconstatacion'] = page.object_list
                    data['fechadesde'] = fechadesde
                    data['fechahasta'] = fechahasta
                    request.session['viewactivo'] = 1
                    return render(request, "af_activofijo/historicoactivoconstatacion.html", data)
                except Exception as ex:
                    messages.error(request, f'{ex}')

            # INFORME BAJA

            elif action == 'addinformebaja':
                try:
                    if 'idex' in request.GET:
                        data['val_extra'] = request.GET['idex']
                    data['id'] = id = encrypt_id(request.GET['id'])
                    activo = ActivoFijo.objects.get(id=id)
                    informebaja = activo.existeinformebaja()
                    if not informebaja:
                        form = InformeBajaFormAF(initial={'solicita': activo.responsable,
                                                          'responsable': persona,
                                                          'estadouso': 2,
                                                          'estado': activo.condicionestado,
                                                          'estadoactivo': activo.estado,
                                                          'enuso': activo.enuso,
                                                          })
                        form.fields['gestion'].queryset = SeccionDepartamento.objects.none()
                        form.fields['solicita'].queryset = Persona.objects.filter(id=activo.responsable.id)
                        form.fields['responsable'].queryset = Persona.objects.filter(id=persona.id)
                    else:
                        informe_dict = model_to_dict(informebaja)
                        form = InformeBajaFormAF(initial=informe_dict)
                        form.fields['solicita'].queryset = Persona.objects.filter(id=informebaja.solicita.id)
                        form.fields['responsable'].queryset = Persona.objects.filter(id=informebaja.responsable.id)
                        data['informe'] = informebaja
                        data['actividades'] = informebaja.actividades_informe_baja()
                    data['switchery'] = True
                    data['form'] = form
                    template = get_template('af_activofijo/modal/forminformebaja.html')
                    return JsonResponse({'result': True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(f'{ex}')
                    return HttpResponseRedirect(request.path)

            elif action == 'verificaciontecnica':
                try:
                    data['id'] = id = encrypt_id(request.GET['id'])
                    data['activo'] = activo = ActivoFijo.objects.get(id=id)
                    form = VerificacionTecnicaForm(initial={
                        'estadoactivo': activo.estado,
                        'condicionestado': activo.condicionestado,
                        'enuso': activo.enuso,
                    })
                    data['switchery'] = True
                    data['form'] = form
                    template = get_template('af_activofijo/modal/formverificaciontecnica.html')
                    return JsonResponse({'result': True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(f'{ex}')
                    return HttpResponseRedirect(request.path)

            elif action == 'historialestadosactivo':
                try:
                    id = int(encrypt(request.GET['id']))
                    data['activo'] = activo = ActivoFijo.objects.get(pk=id)
                    data['historial'] = HistorialEstadoActivo.objects.filter(status=True, activo=activo).exclude(tiporegistro=3).order_by('-id')
                    template = get_template("af_activofijo/modal/historialestadosactivo.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(f'{ex}')
                    return HttpResponseRedirect(request.path)

            elif action == 'addarchivobaja':
                try:
                    activo = ActivoFijo.objects.get(pk=request.GET['id'])
                    if activo.archivobaja:
                        data['filtro'] = filtro = ActivoFijo.objects.get(pk=int(request.GET['id']))
                        data['idactivofijo'] = request.GET['id']
                        data['form2'] = ArchivoActivoBajaForm(initial=model_to_dict(filtro))
                    else:
                        form2 = ArchivoActivoBajaForm()
                        data['filtro'] = filtro = ActivoFijo.objects.get(pk=int(request.GET['id']))
                        data['idactivofijo'] = request.GET['id']
                        data['form2'] = form2
                    template = get_template("af_activofijo/addarchivoactivobaja.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'addinformebajafirmado':
                try:
                    activo = ActivoFijo.objects.get(pk=request.GET['id'])
                    if activo.archivobaja:
                        data['filtro'] = filtro = ActivoFijo.objects.get(pk=int(request.GET['id']))
                        data['idactivofijo'] = request.GET['id']
                        data['form2'] = ArchivoActivoBajaForm(initial=model_to_dict(filtro))
                    else:
                        form2 = ArchivoActivoBajaForm()
                        data['filtro'] = filtro = ActivoFijo.objects.get(pk=int(request.GET['id']))
                        data['idactivofijo'] = request.GET['id']
                        data['form2'] = form2
                    template = get_template("af_activofijo/addarchivoactivobaja.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'firmarinformebaja':
                try:
                    data['val_extra'] = request.GET.get('idex', '')
                    data['id'] = encrypt_id(request.GET['id'])
                    template = get_template("formfirmaelectronicamasiva.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(f'{ex}')
                    return HttpResponseRedirect(request.path)

            elif action == 'firmarinformebajamasivo':
                try:
                    director = get_directorresponsablebaja('AF')
                    es_director = False
                    if director.responsable == persona:
                            data['cond_extra'] = es_director = True
                    if es_director:
                        documentos = DocumentoFirmaInformeBaja.objects.filter(status=True, firmadirector=False, director=director, estadofirma=2, informe__tipoinforme=2)
                    else:
                        documentos = DocumentoFirmaInformeBaja.objects.filter(status=True, firmadirector=False, informe__responsable=persona, estadofirma=1, informe__tipoinforme=2)
                    if not documentos:
                        raise NameError('No tienen documentos pendiente a su nombre para ser firmados.')
                    data['extra_buttons'] = True
                    data['total'] = len(documentos)
                    data['info_mensaje'] = f'Nota: Esta por firmar {len(documentos)} informes de baja que se encuentran a su nombre como responsable.<br>' \
                                              f'Se habilitó la opción de firma por lotes debido a la itermitencia de FIRMA EC'
                    template = get_template("formfirmaelectronicamasiva.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, 'mensaje': f'Error: {ex}'})

            elif action == 'buscarpersonas':
                try:
                    resp = filtro_persona_select(request)
                    return HttpResponse(json.dumps({'status': True, 'results': resp}))
                except Exception as ex:
                    pass

            elif action == 'cargargestion':
                try:
                    departamento=Departamento.objects.get(id=request.GET['id'])
                    resp = [{'value': qs.pk, 'text': f"{qs.descripcion}"}
                            for qs in departamento.gestiones()]
                    return JsonResponse({'result': True, 'data': resp})
                except Exception as ex:
                    pass

            elif action == 'informesbaja':
                try:
                    data['title'] = u'Informes de baja'
                    search, url_vars, estado = request.GET.get('s', ''), f"&action={action}", request.GET.get('estado', '0')
                    mifirma = request.GET.get('mifirma', '0')
                    filtro = Q(status=True, informe__tipoinforme=2)
                    if search:
                        data['s'] = search = request.GET['s']
                        ss = search.split(' ')
                        if len(ss) == 1:
                            filtro = filtro & (Q(informe__activofijo__descripcion=search) |
                                               Q(informe__activofijo__codigogobierno=search) |
                                               Q(informe__activofijo__serie=search) |
                                               Q(informe__activofijo__codigointerno=search) |
                                               Q(informe__activofijo__codigogobierno__icontains=search) |
                                               Q(informe__responsable__nombres__icontains=search) |
                                               Q(informe__responsable__apellido1__icontains=search) |
                                               Q(informe__responsable__apellido2__icontains=search))
                        else:
                            filtro = filtro & (Q(informe__responsable__apellido1__icontains=ss[0]) &
                                               Q(informe__responsable__apellido2__icontains=ss[1]))
                        url_vars += f"&s={search}"

                    if estado != '0':
                        data['estado'] = estado = int(estado)
                        filtro = filtro & Q(estadofirma=estado)
                        url_vars += f"&estado={estado}"

                    documentos = DocumentoFirmaInformeBaja.objects.filter(filtro).order_by('-id')
                    director = get_directorresponsablebaja('AF')

                    if mifirma != '0':
                        data['mifirma'] = mifirma = int(mifirma)
                        url_vars += f"&mifirma={mifirma}"

                        documentos = documentos.filter(director=director)
                        es_director = False
                        if director.responsable == persona:
                            es_director = True

                        if es_director:
                            if mifirma == 1:
                                documentos = documentos.filter(firmadirector=False, estadofirma=2)
                            elif mifirma == 2:
                                documentos = documentos.filter(firmadirector=True)
                        else:
                            documentos = documentos.filter(informe__responsable=persona)
                            if mifirma == 1:
                                documentos = documentos.filter(estadofirma=1)
                            elif mifirma == 2:
                                # list_id_doc = [doc.id for doc in documentos if doc.firmo_responsable_informebaja()]
                                documentos = documentos.filter(estadofirma__in=[2, 3])

                    paging = MiPaginador(documentos, 25)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['listado'] = page.object_list
                    data['url_vars'] = url_vars
                    data['estados'] = MY_ESTADO_FIRMA_INFORME_BAJA
                    data['t_genreado'] = len(documentos.filter(estadofirma=1, status=True))
                    data['t_proceso'] = len(documentos.filter(estadofirma=2, status=True))
                    data['t_legalizado'] = len(documentos.filter(estadofirma=3, status=True))
                    data['t_noactualizado'] = len(documentos.filter(estadofirma=None, status=True))
                    request.session['viewactivo'] = 12
                    return render(request, "af_activofijo/informesbaja.html", data)
                except Exception as ex:
                    messages.error(f'{ex}')
                    return HttpResponseRedirect(request.path)

            elif action == 'historialfirmas':
                try:
                    id = int(encrypt(request.GET['id']))
                    data['documento']=documento=DocumentoFirmaInformeBaja.objects.get(id=id)
                    template = get_template("at_activostecnologicos/modal/historialfirmas.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(f'{ex}')
                    return HttpResponseRedirect(request.path)

            elif action == 'descargarformato':
                try:
                    __author__ = 'Unemi'

                    output = io.BytesIO()
                    workbook = xlsxwriter.Workbook(output)
                    ws = workbook.add_worksheet('formato')
                    formatoceldagris = workbook.add_format(
                        {'align': 'center', 'border': 1, 'text_wrap': True, 'fg_color': '#B6BFC0'})
                    cabeceras= ['codigo gobierno', 'codigo interno',
                                'estado','condicionestado', 'cedula constatador',
                                'responsable activo', 'cedula responsable',
                                'codigoubicacion', 'ubicacion', 'observacion']
                    for index, cabecera in enumerate(cabeceras):
                        ws.write(0, index, cabecera, formatoceldagris)
                    workbook.close()
                    output.seek(0)
                    filename = 'formato_cosntatacion.xlsx'
                    response = HttpResponse(output,
                                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    response['Content-Disposition'] = 'attachment; filename=%s' % filename
                    return response
                except Exception as ex:
                    return JsonResponse({'result': True, 'mensaje': f'Error: {ex}'})

            elif action == 'reportconstatacion':
                try:
                    titulo = 'Generando reporte de constatación de activos.'
                    noti = Notificacion(cuerpo='Reporte de constatación de activos en progreso',
                                        titulo=titulo, destinatario=persona,
                                        url='',
                                        prioridad=1, app_label='sga-sagest',
                                        fecha_hora_visible=datetime.now() + timedelta(days=1), tipo=2,
                                        en_proceso=True)
                    noti.save(request)
                    reporte_activos_constatados_openxl_reportlab_background(request=request, data=data, notify=noti).start()
                    # reporte_activos_constatados_background(request=request, data=data, notif=noti.pk).start()
                    # messages.success(request, 'Generando reporte')
                    return HttpResponseRedirect(f'{request.path}?action=reporteconstatacion&id={request.GET["id_obj"]}')
                except Exception as ex:
                    messages.error(request, f'Error: {ex}')
                    return HttpResponseRedirect(f'{request.path}?action=reporteconstatacion&id={request.GET["id_obj"]}')

            elif action == 'reportconstatacion1':
                try:
                    periodo = PeriodoConstatacionAF.objects.get(id=encrypt_id(request.GET['id_obj']))
                    constatacion, item, responsables, constatadores, estado = request.GET.get('constatacion', ''), \
                                                                              request.GET.get('item', ''), \
                                                                              request.GET.getlist('responsable', ''), \
                                                                              request.GET.getlist('constatador', ''), \
                                                                              request.GET.get('estado', '')
                    filtro, filtro_c, filtro_excl = Q(status=True), Q(status=True), []

                    if estado:
                        filtro = filtro & Q(estado_id=estado)

                    if responsables:
                        filtro = filtro & Q(responsable_id__in=responsables)

                    if constatacion:
                        constatacion = int(constatacion)
                        if constatacion == 1:
                            if constatadores:
                                filtro_c = filtro_c & Q(responsable_id__in=constatadores)
                            if item:
                                if int(item) == 1:
                                    filtro_c = filtro_c & Q(encontraro=True)
                                elif int(item) == 2:
                                    filtro_c = filtro_c & Q(enuso=True)
                                elif int(item) == 3:
                                    filtro_c = filtro_c & Q(requieretraspaso=True)
                                elif int(item) == 4:
                                    filtro_c = filtro_c & Q(requierebaja=True)
                            ids_a = periodo.detalle_constatacion().filter(filtro_c).values_list('activo_id', flat=True)
                            filtro = filtro & Q(id__in=ids_a)
                        elif constatacion == 2:
                            ids_a = periodo.detalle_constatacion().values_list('activo_id', flat=True)
                            filtro_excl = ids_a

                    activos = ActivoFijo.objects.select_related().filter(filtro).exclude(id__in=filtro_excl).order_by('-id')

                    columnas = ['Código Gob',
                                'Código Int',
                                'Activo',
                                'Constatado',
                                'Responsable',
                                'Constatador',
                                'Estado',
                                'En uso',
                                'Encontrado',
                                'Requiere Traspaso',
                                'Requiere dar baja',
                                ]
                    # Crear el DataFrame con los nombres de columna personalizados
                    df = pd.DataFrame(columns=columnas)

                    # Llenar el DataFrame con los datos del queryset
                    for activo in activos:
                        constatado = activo.activo_constatado(periodo)
                        fila = {
                            'Código Gob': activo.codigogobierno,
                            'Código Int': activo.codigointerno,
                            'Activo': activo.descripcion,
                            'Constatado': 'Si' if constatado else 'No',
                            'Responsable': activo.responsable.nombre_completo_minus(),
                            'Constatador': constatado.responsable.nombre_completo_minus() if constatado else '',
                            'Estado': activo.estado,
                            'En uso': 'Si' if constatado and constatado.enuso else 'No',
                            'Encontrado': 'Si' if constatado and constatado.encontrado else 'No',
                            'Requiere Traspaso': 'Si' if constatado and constatado.requieretraspaso else 'No',
                            'Requiere dar baja': 'Si' if constatado and constatado.requieredarbaja else 'No'
                        }
                        fila_df = pd.DataFrame([fila])
                        df = pd.concat([df, fila_df], ignore_index=True)

                    # Crear una respuesta HTTP
                    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    response['Content-Disposition'] = 'attachment; filename=reporte_excel.xlsx'
                    df.to_excel(response, index=False)
                    return response
                except Exception as ex:
                    messages.error(request, f'Error: {ex}')
                    return HttpResponseRedirect(f'{request.path}?action=reporteconstatacion&id={request.GET["id_obj"]}')

            elif action == 'activoscustodiados':
                try:
                    data['title'] = u'Activos custodiados'
                    reporte = request.GET.get('reporte', '')
                    estado, clase, cuenta, search, filtro, url_vars = request.GET.get('estado', ''), \
                                                                      request.GET.get('clase', ''), \
                                                                      request.GET.get('cuenta', ''), request.GET.get('s', ''), Q(status=True, custodio=persona), ''

                    if estado:
                        data['estado'] = int(estado)
                        url_vars += "&estado={}".format(estado)
                        filtro = filtro & Q(estado_id=estado)

                    if clase:
                        data['clase'] = int(clase)
                        url_vars += "&clase={}".format(clase)
                        filtro = filtro & Q(clasebien=clase)

                    if cuenta:
                        data['cuenta'] = int(cuenta)
                        url_vars += "&cuenta={}".format(cuenta)
                        filtro = filtro & Q(cuentacontable=cuenta)

                    if search:
                        ss = search.split(' ')
                        data['s'] = search
                        url_vars += "&s={}".format(search)
                        if len(ss) == 1:
                            filtro = filtro & (Q(codigogobierno=search) |
                                               Q(serie=search) |
                                               Q(codigointerno=search) |
                                               Q(responsable__cedula=search) |
                                               Q(responsable__pasaporte=search) |
                                               Q(responsable__nombres__icontains=search) |
                                               Q(responsable__apellido1__icontains=search) |
                                               Q(responsable__apellido2__icontains=search))
                        else:
                            filtro = filtro & (Q(responsable__apellido1__icontains=ss[0]) &
                                               Q(responsable__apellido2__icontains=ss[1]) |
                                               Q(descripcion=search))

                    activos = ActivoFijo.objects.filter(filtro).order_by('-fechaingreso').distinct()
                    if reporte:
                        if reporte == 'pdf':
                            data = {}
                            data['activos'] = activos
                            data['hoy'] = datetime.now()
                            data['custodio'] = activos.first().custodio
                            context = {'pagesize': 'A4 landscape', 'data': data}
                            dir = 'af_activofijo/informes/reporte_activos_fijos.html'
                            return conviert_html_to_pdf(dir, context)
                        else:
                            return reporte_activos(activos)
                    else:
                        paging = MiPaginador(activos, 25)
                        p = 1
                        try:
                            paginasesion = 1
                            if 'paginador' in request.session:
                                paginasesion = int(request.session['paginador'])
                            if 'page' in request.GET:
                                p = int(request.GET['page'])
                            else:
                                p = paginasesion
                            try:
                                page = paging.page(p)
                            except:
                                p = 1
                            page = paging.page(p)
                        except:
                            page = paging.page(p)
                        request.session['paginador'] = p
                        data['paging'] = paging
                        data['rangospaging'] = paging.rangos_paginado(p)
                        data['usuario'] = usuario
                        data['page'] = page
                        data['activos'] = page.object_list
                        data['estados'] = EstadoProducto.objects.filter(status=True)
                        data['clases'] = CLASE_BIEN
                        data['url_vars'] = url_vars
                        data['cuentas'] = CuentaContable.objects.filter(status=True, activosfijos=True)
                        data['fecha'] = datetime.now().date()
                        data['total'] = len(activos)
                        request.session['viewactivo'] = 13
                        return render(request, "af_activofijo/activoscustodiados.html", data)
                except Exception as ex:
                    messages.error(request, f'Error {ex}')

            elif action == 'constatacionesusuario':
                try:
                    data['title'] = u'Constataciones de activos por ususarios'
                    data['periodo_c'] = periodo = PeriodoConstatacionAF.objects.get(id=int(encrypt(request.GET['id'])))
                    constatador, estado,estadoacta, search, url_vars, filtro = request.GET.get('constatador', ''), \
                                                                                   request.GET.get('estado', ''), \
                                                                                   request.GET.get('estadoacta', ''), \
                                                                                   request.GET.get('s', ''), \
                                                                                   f'&action={action}&id={request.GET["id"]}', \
                                                                                   Q(status=True, periodo=periodo)
                    if estado:
                        data['estado'] = int(estado)
                        url_vars += "&estado={}".format(estado)
                        filtro = filtro & Q(estado=estado)
                    if estadoacta:
                        data['estadoacta'] = estadoacta = int(estadoacta)
                        filtro = filtro & Q(estadoacta=estadoacta)
                        url_vars += "&estadoaaacta={}".format(estadoacta)

                    if constatador:
                        data['constatador']=constatador=int(constatador)
                        url_vars += "&constatador={}".format(constatador)
                        ids_a=periodo.detalle_constatacion().filter(responsable_id=constatador).values_list('activo_id',flat=True)
                        filtro = filtro & Q(id__in=ids_a)

                    if search:
                        ss = search.split(' ')
                        data['s'] = search
                        url_vars += "&s={}".format(search)
                        if len(ss) == 1:
                            filtro = filtro & (
                                               Q(usuariobienes__cedula=search) |
                                               Q(usuariobienes__pasaporte=search) |
                                               Q(usuariobienes__apellido1__icontains=search) |
                                               Q(usuariobienes__apellido2__icontains=search) |
                                               Q(usuariobienes__nombres__icontains=search))
                        else:
                            filtro = filtro & (Q(usuariobienes__apellido1__icontains=ss[0]) &
                                               Q(usuariobienes__apellido2__icontains=ss[1]) |
                                               Q(usuariobienes__nombres__icontains=search))

                    constataciones = ConstatacionFisica.objects.filter(filtro).order_by('estado', '-numero')
                    paging = MiPaginador(constataciones, 10)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['constatadores']=periodo.constatadores()
                    data['estados'] = ESTADO_CONSTATACION
                    data['estadosacta'] = ESTADO_ACTA_CONSTATACION
                    data['reporte_0'] = obtener_reporte('consulta_activos_catalogo')
                    data['url_vars'] = url_vars
                    data['listado'] = page.object_list
                    request.session['viewactivo'] = 2
                    return render(request, "af_activofijo/constataciones_usuarios.html", data)
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'descargaractasconstatacion':
                try:
                    data['form'] = DescargarActasConstatacionForm()
                    data['id'] = request.GET['id']
                    template = get_template('af_activofijo/modal/formdescargaractasconstatacion.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'det_constatacion':
                try:
                    data['constatacion'] = constatacion = ConstatacionFisica.objects.get(pk=encrypt_id(request.GET['id']))
                    data['activos_constatados'] = activos_constatados = constatacion.detalle_constatacion()
                    data['activos_faltantes'] = ActivoFijo.objects.filter(status=True, responsable=constatacion.usuariobienes, statusactivo=1).exclude(id__in=activos_constatados.values_list('activo_id', flat=True))
                    template = get_template("af_activofijo/modal/detalleconstatacionesusuario.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": True, 'data': json_content})
                except Exception as ex:
                    return JsonResponse({"result": False, "mensaje": u"Error al obtener los datos."})

            elif action == 'actaconstatacion':
                try:
                    data['tipo'] = tipo
                    data['fechahoy'] = datetime.now().date()
                    directory = os.path.join(MEDIA_ROOT, 'reportes', 'crai')
                    valido = True
                    nombre_archivo = 'acta_constatacion' + str(random.randint(1, 10000)) + '.pdf'
                    try:
                        os.stat(directory)
                    except:
                        os.mkdir(directory)
                    valido = conviert_html_to_pdfsave_generic_lotes(
                        request,
                        'af_activofijo/informes/acta_constatacion.html',
                        {
                            'pagesize': 'A4 landscape',
                            'data': data,
                        },
                        directory, nombre_archivo
                    )
                except Exception as ex:
                    pass

            elif action == 'firmaractaconstatacion':
                try:
                    data['id'] = id = int(encrypt(request.GET['id']))
                    qr = qrImgFirma(request, persona, "png", paraMostrar=True)
                    data["qrBase64"] = qr[0]
                    data['filtro'] = filtro = ConstatacionFisica.objects.get(id=id)
                    data['archivo_url'] = filtro.get_documento().url
                    template = get_template("formfirmaelectronica_v2.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'firmarconstatacionmasivo':
                try:
                    data['val_extra'] = request.GET.get('idex', '')
                    data['id'] = id = encrypt_id(request.GET['id'])
                    director = SeccionDepartamento.objects.get(id=23).responsable
                    actas_firmadas = ActaConstatacion.objects.filter(status=True,constatacion__periodo_id=id, persona=persona, estado=2).values_list('constatacion_id', flat=True).order_by('constatacion_id').distinct()
                    if persona == director:
                        constataciones = ConstatacionFisica.objects.filter(status=True, estado=2, periodo_id=id, estadoacta=2).exclude(id__in=actas_firmadas)
                    else:
                        constataciones = ConstatacionFisica.objects.filter(status=True, usuariofinaliza=persona, estado=2, periodo_id=id, estadoacta=1).exclude(id__in=actas_firmadas)

                    if not constataciones:
                        raise NameError('No tienen documentos pendiente a su nombre por ser firmados.')
                    data['info_mensaje'] = f'Esta por firmar {len(constataciones)} actas de constatación que se encuentran a su nombre.'
                    template = get_template("formfirmaelectronicamasiva.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, 'mensaje': f'Error: {ex}'})

            elif action == 'directorresponsablebaja':
                try:
                    data['title'] = u'Responsable informe de baja'
                    reporte = request.GET.get('reporte', '')
                    estado, clase, cuenta, search, filtro, url_vars = request.GET.get('estado', ''), \
                                                                      request.GET.get('clase', ''), \
                                                                      request.GET.get('cuenta', ''), request.GET.get('s', ''), Q(status=True), f'&action={action}'


                    if search:
                        ss = search.split(' ')
                        data['s'] = search
                        url_vars += "&s={}".format(search)
                        if len(ss) == 1:
                            filtro = filtro & (Q(responsable__cedula=search) |
                                               Q(responsable__pasaporte=search) |
                                               Q(responsable__nombres__icontains=search) |
                                               Q(responsable__apellido1__icontains=search) |
                                               Q(responsable__apellido2__icontains=search))
                        else:
                            filtro = filtro & (Q(responsable__apellido1__icontains=ss[0]) &
                                               Q(responsable__apellido2__icontains=ss[1]))
                        url_vars += f"&s={search}"

                    directores = DirectorResponsableBaja.objects.filter(filtro).order_by('-id')
                    paging = MiPaginador(directores, 15)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['usuario'] = usuario
                    data['page'] = page
                    data['directores'] = page.object_list
                    data['url_vars'] = url_vars
                    data['fecha'] = datetime.now().date()
                    request.session['viewactivo'] = 14
                    return render(request, "af_activofijo/directorresponsablebaja.html", data)

                except Exception as ex:
                    messages.error(request, f'Error {ex}')

            elif action == 'adddirectorresponsablebaja':
                try:
                    form = ResponsableInformeBajaForm()
                    form.fields['responsable'].queryset = Persona.objects.none()
                    id = request.GET.get('id', '')
                    if id:
                        data['id'] = id = encrypt_id(id)
                        director = DirectorResponsableBaja.objects.get(pk=id)
                        form.initial = model_to_dict(director)
                        form.fields['responsable'].queryset = Persona.objects.filter(id=director.responsable.id)
                    data['switchery'] = True
                    data['form'] = form
                    template = get_template('af_activofijo/modal/formdirectorresponsablebaja.html')
                    return JsonResponse({'result': True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(f'{ex}')
                    return HttpResponseRedirect(request.path)

            return HttpResponseRedirect(request.path)

        else:
            data['title'] = u'Activos y bienes de la institución'
            reporte = request.GET.get('reporte', '')
            ids = None
            estado = request.GET.get('estado', '')
            clase = request.GET.get('clase', '')
            cuenta = request.GET.get('cuenta', '')
            search = request.GET.get('s', '')
            filtro = Q(status=True)
            url_vars = ''
            responsable = request.GET.get('responsable', '')
            ubicacion = request.GET.get('ubicacion', '')

            if estado:
                data['estado'] = int(estado)
                url_vars += "&estado={}".format(estado)
                filtro = filtro & Q(estado_id=estado)

            if clase:
                data['clase'] = int(clase)
                url_vars += "&clase={}".format(clase)
                filtro = filtro & Q(clasebien=clase)

            if cuenta:
                data['cuenta'] = int(cuenta)
                url_vars += "&cuenta={}".format(cuenta)
                filtro = filtro & Q(cuentacontable=cuenta)

            if search:
                ss = search.split(' ')
                data['s'] = search
                url_vars += "&s={}".format(search)
                if len(ss) == 1:
                    filtro = filtro & (Q(codigogobierno=search) |
                                       Q(serie=search) |
                                       Q(codigointerno=search) |
                                       Q(responsable__cedula=search) |
                                       Q(responsable__pasaporte=search) |
                                       Q(responsable__nombres__icontains=search) |
                                       Q(responsable__apellido1__icontains=search) |
                                       Q(responsable__apellido2__icontains=search))
                else:
                    filtro = filtro & (Q(responsable__apellido1__icontains=ss[0]) &
                                       Q(responsable__apellido2__icontains=ss[1]) |
                                       Q(descripcion=search))

            if responsable:
                persona = Persona.objects.get(pk=encrypt_id(responsable))
                data['responsable'] = {'id': persona.id, 'nombre': persona.nombre_completo_inverso()}
                url_vars += "&responsable={}".format(responsable)
                filtro = filtro & Q(responsable_id=responsable)

            if ubicacion:
                data['ubicacion'] = int(ubicacion)
                url_vars += "&ubicacion={}".format(ubicacion)
                filtro = filtro & Q(ubicacion_id=ubicacion)

            if 'id' in request.GET:
                ids = int(request.GET['id'])
                filtro = Q(status=True)

            activos = ActivoFijo.objects.filter(filtro).order_by('-fechaingreso').distinct()
            if reporte:
                return reporte_activos(activos)
            else:
                paging = MiPaginador(activos, 25)
                p = 1
                try:
                    paginasesion = 1
                    if 'paginador' in request.session:
                        paginasesion = int(request.session['paginador'])
                    if 'page' in request.GET:
                        p = int(request.GET['page'])
                    else:
                        p = paginasesion
                    try:
                        page = paging.page(p)
                    except:
                        p = 1
                    page = paging.page(p)
                except:
                    page = paging.page(p)
                request.session['paginador'] = p
                data['paging'] = paging
                data['rangospaging'] = paging.rangos_paginado(p)
                data['usuario'] = usuario
                data['page'] = page
                data['search'] = search if search else ""
                data['ids'] = ids if ids else ""
                data['activos'] = page.object_list
                data['estados'] = EstadoProducto.objects.filter(status=True).exclude(id__in=[4, 5]).order_by('id')
                data['ubicaciones'] = Ubicacion.objects.values_list('id', 'nombre', 'bloque__descripcion').filter(status=True)
                data['clases'] = CLASE_BIEN
                data['url_vars'] = url_vars
                data['cuentas'] = CuentaContable.objects.filter(status=True, activosfijos=True)
                data['fecha'] = datetime.now().date()
                request.session['viewactivo'] = 1
                return render(request, "af_activofijo/view.html", data)


def reporte_activos(activos):
    try:
        __author__ = 'Unemi'
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output)
        ws = workbook.add_worksheet('Listado')
        ws.set_column(0, 2, 20)
        ws.set_column(3, 23, 60)

        formatoceldagris = workbook.add_format(
            {'align': 'center', 'border': 1, 'text_wrap': True, 'fg_color': '#B6BFC0'})

        ws.write(0, 0, 'Nro.', formatoceldagris)
        ws.write(0, 1, 'Cod. gobierno', formatoceldagris)
        ws.write(0, 2, 'Cod. interno', formatoceldagris)
        ws.write(0, 3, 'Fecha ingreso', formatoceldagris)
        ws.write(0, 4, 'Catalogo activo', formatoceldagris)
        ws.write(0, 5, 'Descripción', formatoceldagris)
        ws.write(0, 6, 'Marca', formatoceldagris)
        ws.write(0, 7, 'Modelo', formatoceldagris)
        ws.write(0, 8, 'Serie', formatoceldagris)
        ws.write(0, 9, 'Ubicación', formatoceldagris)
        ws.write(0, 10, 'Responsable', formatoceldagris)
        ws.write(0, 11, 'V.U.', formatoceldagris)
        ws.write(0, 12, 'Cuenta contable.', formatoceldagris)
        ws.write(0, 13, 'Costo', formatoceldagris)
        ws.write(0, 14, 'Valor residual', formatoceldagris)
        ws.write(0, 15, 'Valor libros', formatoceldagris)
        ws.write(0, 16, 'Dep. acumulada', formatoceldagris)
        ws.write(0, 17, 'Estado', formatoceldagris)
        ws.write(0, 18, 'Número', formatoceldagris)
        ws.write(0, 19, 'Observación', formatoceldagris)
        ws.write(0, 20, 'Clase bien', formatoceldagris)
        ws.write(0, 21, 'Estado de baja', formatoceldagris)
        ws.write(0, 22, 'Subido al gobierno', formatoceldagris)
        ws.write(0, 23, 'Cargo de responsable', formatoceldagris)
        ws.write(0, 24, 'Archivo de Baja', formatoceldagris)
        cont = 1

        for activo in activos:
            cargo = DistributivoPersona.objects.filter(status=True, persona=activo.responsable).first()
            ws.write(cont, 0, str(cont))
            ws.write(cont, 1, str(activo.codigogobierno))
            ws.write(cont, 2, str(activo.codigointerno))
            ws.write(cont, 3, str(activo.fechacomprobante))
            ws.write(cont, 4, str(activo.catalogo))
            ws.write(cont, 5, str(activo.descripcion))
            ws.write(cont, 6, str(activo.marca))
            ws.write(cont, 7, str(activo.modelo))
            ws.write(cont, 8, str(activo.serie))
            ws.write(cont, 9, str(activo.ubicacion))
            ws.write(cont, 10, str(activo.responsable))
            ws.write(cont, 11, str(activo.vidautil))
            ws.write(cont, 12, str(activo.cuentacontable))
            ws.write(cont, 13, str(activo.costo))
            ws.write(cont, 14, str(activo.valorresidual))
            ws.write(cont, 15, str(activo.valorlibros))
            ws.write(cont, 16, str(activo.valordepreciacionacumulada))
            ws.write(cont, 17, str(activo.estado))
            ws.write(cont, 18, str(activo.numerocomprobante))
            ws.write(cont, 19, str(activo.observacion))
            ws.write(cont, 20, str(activo.get_clasebien_display()))
            ws.write(cont, 21, str(activo.get_statusactivo_display()))
            ws.write(cont, 22, "SI" if activo.subidogobierno else "NO")
            ws.write(cont, 23, str(cargo.denominacionpuesto if cargo else ''))
            ws.write(cont, 24, "Si" if activo.archivobaja else "No")
            cont += 1
        workbook.close()
        output.seek(0)
        filename = 'reporte_activos.xlsx'
        response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=%s' % filename
        return response
    except Exception as ex:
        raise NameError(f'{ex}')

def crear_constataciones(activo, persona, request):
    with transaction.atomic():
        try:
            hoy = datetime.now()
            periodo = PeriodoConstatacionAF.objects.filter(status=True, activo=True, cerrado=False, fechainicio__lte=hoy, fechafin__gte=hoy).first()
            if periodo:
                constatacion = ConstatacionFisica.objects.filter(usuariobienes=activo.responsable, periodo=periodo).first()
                if not constatacion:
                    secuencia = secuencia_activos(request)
                    secuencia.numeroconstatacion += 1
                    secuencia.save(request)
                    constatacion = ConstatacionFisica(usuariobienes=activo.responsable,
                                                      numero=secuencia.numeroconstatacion,
                                                      normativaconstatacion=secuencia.normativaconstatacion,
                                                      fechainicio=hoy,
                                                      periodo=periodo,
                                                      ubicacionbienes=activo.ubicacion)
                    constatacion.save(request)
                    log(u'Agrego constatación: %s' % (constatacion), request, "add")

                detalle_c = DetalleConstatacionFisica(codigoconstatacion=constatacion,
                                                      activo=activo,
                                                      responsable=persona,
                                                      usuariobienes=activo.responsable,
                                                      ubicacionbienes=activo.ubicacion,
                                                      estadooriginal=activo.estado,
                                                      estadoactual=activo.estado,
                                                      enuso=True,
                                                      perteneceusuario = True,
                                                      encontrado=True)
                detalle_c.save(request)
                log(u'Agrego detalle de constatación: %s' % (detalle_c), request, "add")
        except Exception as ex:
            transaction.set_rollback(True)
            raise NameError(f'Error: {ex}')

def director_af(informebaja):
    director = PersonaDepartamentoFirmas.objects.filter(status=True, tipopersonadepartamento_id=1, departamentofirma_id=2)
    if director.filter(fechafin__gte=informebaja.fecha_creacion.date(), fechainicio__lte=informebaja.fecha_creacion.date()):
        director = director.filter(fechafin__gte=informebaja.fecha_creacion.date(), fechainicio__lte=informebaja.fecha_creacion.date()).first()
    elif director.filter(actualidad=True):
        director = director.filter(actualidad=True).first()
    return director

def get_directorresponsablebaja(departamento):
    grupo_id = 12 if departamento == 'AF' else 14
    grupo = GrupoDepartamento.objects.filter(id=grupo_id).first()
    if not grupo:
        raise NameError('No existe un grupo de departamento.')
    departamento = grupo.departamento_set.filter(status=True, integrantes__isnull=False).first()
    if not departamento:
        raise NameError('No existe un departamento.')
    directorresponsable = DirectorResponsableBaja.objects.filter(status=True, departamento=departamento, actual=True).first()
    if not directorresponsable:
        raise NameError('No existe un director responsable vigente.')
    return directorresponsable

def get_ruta_hoja_membretada():
    ruta_hoja_membretada = os.path.join(SITE_STORAGE, 'media', 'reportes', 'encabezados_pies', 'hoja_membretada.png')
    ruta_hoja_membretada.replace('\\', '/')
    return ruta_hoja_membretada

def generar_informebaja(request, informebaja):
    try:
        # director = director_af(informebaja)
        directorresponsable = get_directorresponsablebaja('AF')
        # if not director:
        #     raise NameError('No existe un director vigente.')
        if not informebaja:
            raise NameError('No existe informe de baja creado.')

        perso = informebaja.responsable
        detalle = informebaja.detalleinformeactivobaja_set.filter(status=True)
        directory = os.path.join(MEDIA_ROOT, 'activos', 'informesbaja')
        nombre_archivo = f'Informe_Baja{informebaja.activofijo.id}' + '.pdf'
        os.makedirs(directory, exist_ok=True)
        ruta_hoja_membretada = get_ruta_hoja_membretada()
        dir = 'af_activofijo/informebajapdfmantenimiento.html'
        valido = conviert_html_to_pdfsaveqr_generico(request, dir,
                                                     {'pagesize': 'A4', 'ruta_hoja_membretada': ruta_hoja_membretada,
                                                      'informe': informebaja.activofijo, 'perso': perso, 'director': directorresponsable,
                                                      'bajaactivo': informebaja, 'detalle': detalle, 'hoy': datetime.now().date()
                                                      }, directory, nombre_archivo)
        if not valido[0]:
            raise NameError('Error al generar el informe')
        archivo_url = '/media/activos/informesbaja/' + nombre_archivo
        ruta_archivo = os.path.join(SITE_STORAGE, 'media', 'activos', 'informesbaja', nombre_archivo)

        # Obtiene el contenido del archivo desde la ubicación de medios
        with default_storage.open(ruta_archivo, "rb") as pdf_file:
            # Crea un objeto de archivo Django a partir del contenido
            pdf_content = ContentFile(pdf_file.read())
            pdf_file = DjangoFile(pdf_content)
        return pdf_file, archivo_url
    except Exception as ex:
        raise NameError(f'Error: {ex}')

def generar_informebaja_pdf_v2(informe, director):
    try:
        ruta_hoja_membretada = get_ruta_hoja_membretada()
        perso = informe.responsable
        detalle = informe.detalleinformeactivobaja_set.filter(status=True)
        nombre_archivo = generar_nombre(f'Informe_Baja{informe.activofijo.id}_', 'generado') + '.pdf'
        template_path = 'af_activofijo/informebajapdfmantenimiento.html'
        context = {'pagesize': 'A4', 'ruta_hoja_membretada': ruta_hoja_membretada,
                   'informe': informe.activofijo, 'perso': perso, 'director': director,
                   'bajaactivo': informe, 'detalle': detalle, 'hoy': datetime.now().date()
                   }
        file_obj, response = conviert_html_to_pdf_save_file_model(template_path, context, nombre_archivo)
        return file_obj, response
    except Exception as ex:
        raise NameError(f'Error: {ex}')

def firmar_doc_informebaja(request, documento):
    try:
        informebaja = documento.informe
        persona = informebaja.responsable
        certificado = request.FILES["firma"]
        contrasenaCertificado = request.POST['palabraclave']
        extension_certificado = os.path.splitext(certificado.name)[1][1:]
        bytes_certificado = certificado.read()
        archivo_url = documento.archivo.url
        archivo_ = documento.archivo
        cargo = persona.mi_cargo_administrativo() if persona.mi_cargo_administrativo() else ''
        palabras = f'{persona} {cargo}'
        x, y, numPage = obtener_posicion_x_y_saltolinea(archivo_url, palabras, False, True)
        if not x or not y:
            raise NameError('No se encontró el responsable en el documento.')
        datau = JavaFirmaEc(archivo_a_firmar=archivo_, archivo_certificado=bytes_certificado,
                            extension_certificado=extension_certificado,
                            password_certificado=contrasenaCertificado,
                            page=int(numPage), reason='', lx=x, ly=y).sign_and_get_content_bytes()
        archivo_ = io.BytesIO()
        archivo_.write(datau)
        archivo_.seek(0)

        _name = f"InformeBaja_{informebaja.id}"
        file_obj = DjangoFile(archivo_, name=f"{_name}.pdf")
        return file_obj
    except Exception as ex:
        raise NameError(f'Error: {ex}')


