#!/usr/bin/env python
# -*- coding: utf-8 -*-
import os
import sys
import openpyxl

# SITE_ROOT = os.path.dirname(os.path.realpath(__file__))
YOUR_PATH = os.path.dirname(os.path.realpath(__file__))
# print(f"YOUR_PATH: {YOUR_PATH}")
SITE_ROOT = os.path.dirname(os.path.dirname(YOUR_PATH))
SITE_ROOT = os.path.join(SITE_ROOT, '')
# print(f"SITE_ROOT: {SITE_ROOT}")
your_djangoproject_home = os.path.split(SITE_ROOT)[0]
# print(f"your_djangoproject_home: {your_djangoproject_home}")
sys.path.append(your_djangoproject_home)

from django.core.wsgi import get_wsgi_application
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "settings")
application = get_wsgi_application()
from sga.models import *
from sagest.models import *
from sga.funciones import log, salvaRubros
# from balcon.models import *
from moodle import moodle
from certi.models import *
from posgrado.models import InscripcionCohorte, AsesorComercial, HistorialAsesor, \
    EvidenciaRequisitosAspirante, RequisitosMaestria, CohorteMaestria, ConfigFinanciamientoCohorte, \
    Contrato, DetalleEvidenciaRequisitosAspirante, \
    DetalleAprobacionContrato, DetalleAprobacionFormaPago, VentasProgramaMaestria, MaestriasAdmision, \
    Requisito, AsesorMeta, CambioAdmitidoCohorteInscripcion, RespuestaEvaluacionAcreditacionPosgrado, \
    RespuestaRubricaPosgrado, DetalleRespuestaRubricaPos, DetalleRespuestaRubricaPosgrado
from sga.commonviews import ficha_socioeconomica
from Moodle_Funciones import crearhtmlphpmoodle
from inno.models import TipoFormaPagoPac
from settings import MEDIA_ROOT, MEDIA_URL
import xlsxwriter
import warnings
from django.http import HttpResponse
from django.db.models.query_utils import Q
import io
warnings.filterwarnings('ignore', message='Unverified HTTPS request')
print(u"********************Arreglo migración de datos de evaluación***************************")

try:
    filtro = Q(status=True, materia__nivel__periodo__tipo__id=3, tipoprofesor__id=11,
               materia__fin__lt=datetime.now().date(), materia__fin__range=('2024-07-01', '2024-09-05'))
    # filtro = Q(status=True, materia__nivel__periodo__tipo__id=3, tipoprofesor__id=11, materia__id=100374,
    #            materia__fin__lt=datetime.now().date())

    eProfesoresMateria = ProfesorMateria.objects.filter(filtro).order_by('-id')
    cordi = Coordinacion.objects.get(pk=7)
    c = 1
    for profesormateria in eProfesoresMateria:
        proceso = profesormateria.materia.nivel.periodo.proceso_evaluativo()
        materia = profesormateria.materia
        print(f'{materia}')
        print("******HETERO*****")
        if RespuestaEvaluacionAcreditacion.objects.filter(proceso=proceso, tipoinstrumento=1, profesor=profesormateria.profesor, tipoprofesor=profesormateria.tipoprofesor,
                                            materia=materia,
                                            coordinacion=materia.coordinacion_materia(),
                                            carrera=materia.carrera()).exists():

            eRespuestas = RespuestaEvaluacionAcreditacion.objects.filter(proceso=proceso, tipoinstrumento=1,
                                                                         profesor=profesormateria.profesor,
                                                                         tipoprofesor=profesormateria.tipoprofesor,
                                                                         materia=materia,
                                                                         coordinacion=materia.coordinacion_materia(),
                                                                         carrera=materia.carrera())

            cvnt = 0
            for eRespuesta in eRespuestas:
                if not RespuestaEvaluacionAcreditacionPosgrado.objects.filter(proceso__periodo=eRespuesta.proceso.periodo,
                                                                          profesor=eRespuesta.profesor,
                                                                          evaluador=eRespuesta.evaluador, materia=eRespuesta.materia,
                                                                          tipoinstrumento=1).exists():
                    eRespuestaRubrica = RespuestaRubrica.objects.filter(status=True, respuestaevaluacion=eRespuesta, rubrica__para_hetero=True).order_by('-id').first()
                    eDetallesRespuestaRubrica = DetalleRespuestaRubrica.objects.filter(respuestarubrica=eRespuestaRubrica).order_by('id')

                    evaluacion = RespuestaEvaluacionAcreditacionPosgrado(proceso=eRespuesta.proceso,
                                                                         tipoinstrumento=1,
                                                                         profesor=eRespuesta.profesor,
                                                                         tipoprofesor=eRespuesta.tipoprofesor if eRespuesta.tipoprofesor else None,
                                                                         evaluador=eRespuesta.evaluador if eRespuesta.evaluador else None,
                                                                         materia=eRespuesta.materia,
                                                                         materiaasignada=eRespuesta.materiaasignada if eRespuesta.materiaasignada else None,
                                                                         fecha=eRespuesta.fecha,
                                                                         coordinacion=eRespuesta.coordinacion if eRespuesta.coordinacion else cordi,
                                                                         accionmejoras=eRespuesta.accionmejoras,
                                                                         formacioncontinua=eRespuesta.formacioncontinua,
                                                                         carrera=eRespuesta.carrera if eRespuesta.carrera else None)
                    evaluacion.save()

                    print(f"Cabecera respuesta creada - evaluador {evaluacion.evaluador}")
                    if not RespuestaRubricaPosgrado.objects.filter(respuestaevaluacion=evaluacion).exists():
                        respuestarubrica = RespuestaRubricaPosgrado(respuestaevaluacion=evaluacion,
                                                                    rubrica=eRespuestaRubrica.rubrica,
                                                                    valor=0)
                        respuestarubrica.save()
                    else:
                        respuestarubrica = RespuestaRubricaPosgrado.objects.filter(respuestaevaluacion=evaluacion, rubrica=eRespuestaRubrica.rubrica)[0]
                    print(f"Cabecera rubrica creada - rubrica {respuestarubrica.rubrica}")

                    for eDetalle in eDetallesRespuestaRubrica:
                        eDetalleNew = DetalleRespuestaRubricaPos(respuestarubrica=respuestarubrica,
                                                                 rubricapregunta=eDetalle.rubricapregunta,
                                                                 valor=eDetalle.valor)
                        eDetalleNew.save()
                        respuestarubrica.save()

                    print(f"Respuestas migradas {eDetallesRespuestaRubrica.count()}")
                    cvnt += 1
                print(f'{cvnt}/{eRespuestas.count()}')

        print("******AUTO*****")
        if RespuestaEvaluacionAcreditacion.objects.filter(proceso=proceso, tipoinstrumento=2, profesor=profesormateria.profesor,
                                            materia=materia).exists():

            eRespuestas = RespuestaEvaluacionAcreditacion.objects.filter(proceso=proceso,
                                                                         tipoinstrumento=2,
                                                                         profesor=profesormateria.profesor,
                                                                         materia=materia)

            cvnt = 0
            for eRespuesta in eRespuestas:
                if not RespuestaEvaluacionAcreditacionPosgrado.objects.filter(proceso__periodo=eRespuesta.proceso.periodo,
                                                                              profesor=eRespuesta.profesor, tipoinstrumento=2,
                                                                              materia=eRespuesta.materia).exists():
                    eRespuestaRubrica = RespuestaRubrica.objects.filter(status=True, respuestaevaluacion=eRespuesta, rubrica__para_auto=True).order_by('-id').first()
                    eDetallesRespuestaRubrica = DetalleRespuestaRubricaPosgrado.objects.filter(respuestarubrica=eRespuestaRubrica).order_by('id')

                    evaluacion = RespuestaEvaluacionAcreditacionPosgrado(proceso=eRespuesta.proceso,
                                                                         tipoinstrumento=2,
                                                                         profesor=eRespuesta.profesor,
                                                                         fecha=eRespuesta.fecha,
                                                                         coordinacion=eRespuesta.coordinacion,
                                                                         carrera=eRespuesta.carrera,
                                                                         accionmejoras=eRespuesta.accionmejoras,
                                                                         tipomejoras=eRespuesta.tipomejoras,
                                                                         materia=eRespuesta.materia)
                    evaluacion.save()

                    print(f"Cabecera respuesta creada - evaluador {evaluacion.evaluador}")
                    if not RespuestaRubricaPosgrado.objects.filter(respuestaevaluacion=evaluacion).exists():
                        respuestarubrica = RespuestaRubricaPosgrado(respuestaevaluacion=evaluacion,
                                                                    rubrica=eRespuestaRubrica.rubrica,
                                                                    valor=0)
                        respuestarubrica.save()
                    else:
                        respuestarubrica = RespuestaRubricaPosgrado.objects.filter(respuestaevaluacion=evaluacion, rubrica=eRespuestaRubrica.rubrica)[0]
                    print(f"Cabecera rubrica creada - rubrica {respuestarubrica.rubrica}")

                    for eDetalle in eDetallesRespuestaRubrica:
                        eDetalleNew = DetalleRespuestaRubricaPos(respuestarubrica=respuestarubrica,
                                                                 rubricapregunta=eDetalle.rubricapregunta,
                                                                 valor=eDetalle.valor)
                        eDetalleNew.save()
                        respuestarubrica.save()

                    print(f"Respuestas migradas {eDetallesRespuestaRubrica.count()}")
                    cvnt += 1
                print(f'{cvnt}/{eRespuestas.count()}')

        print("******DIRECTIVOS*****")

        if RespuestaEvaluacionAcreditacion.objects.filter(proceso=proceso, tipoinstrumento=4, profesor=profesormateria.profesor,
                                                            materia=materia).exists():

            eRespuestas = RespuestaEvaluacionAcreditacion.objects.filter(proceso=proceso, tipoinstrumento=4,
                                                                         profesor=profesormateria.profesor,
                                                                         materia=materia)

            cvnt = 0
            for eRespuesta in eRespuestas:
                if not RespuestaEvaluacionAcreditacionPosgrado.objects.filter(proceso__periodo=eRespuesta.proceso.periodo,
                                                                          profesor=eRespuesta.profesor,
                                                                          evaluador=eRespuesta.evaluador, materia=eRespuesta.materia,
                                                                          tipoinstrumento=4).exists():
                    eRespuestaRubrica = RespuestaRubrica.objects.filter(status=True, respuestaevaluacion=eRespuesta, rubrica__para_directivo=True).order_by('-id').first()
                    eDetallesRespuestaRubrica = DetalleRespuestaRubricaPosgrado.objects.filter(respuestarubrica=eRespuestaRubrica).order_by('id')

                    evaluacion = RespuestaEvaluacionAcreditacionPosgrado(proceso=eRespuesta.proceso,
                                                                         tipoinstrumento=4,
                                                                         profesor=eRespuesta.profesor,
                                                                         evaluador=eRespuesta.evaluador,
                                                                         fecha=eRespuesta.fecha,
                                                                         coordinacion=eRespuesta.coordinacion,
                                                                         carrera=eRespuesta.carrera,
                                                                         tipomejoras=eRespuesta.tipomejoras,
                                                                         accionmejoras=eRespuesta.accionmejoras,
                                                                         materia=eRespuesta.materia)
                    evaluacion.save()

                    print(f"Cabecera respuesta creada - evaluador {evaluacion.evaluador}")
                    if not RespuestaRubricaPosgrado.objects.filter(respuestaevaluacion=evaluacion).exists():
                        respuestarubrica = RespuestaRubricaPosgrado(respuestaevaluacion=evaluacion,
                                                                    rubrica=eRespuestaRubrica.rubrica,
                                                                    valor=0)
                        respuestarubrica.save()
                    else:
                        respuestarubrica = RespuestaRubricaPosgrado.objects.filter(respuestaevaluacion=evaluacion, rubrica=eRespuestaRubrica.rubrica)[0]
                    print(f"Cabecera rubrica creada - rubrica {respuestarubrica.rubrica}")

                    for eDetalle in eDetallesRespuestaRubrica:
                        eDetalleNew = DetalleRespuestaRubricaPos(respuestarubrica=respuestarubrica,
                                                                 rubricapregunta=eDetalle.rubricapregunta,
                                                                 valor=eDetalle.valor)
                        eDetalleNew.save()
                        respuestarubrica.save()

                    print(f"Respuestas migradas {eDetallesRespuestaRubrica.count()}")
                    cvnt += 1
                print(f'{cvnt}/{eRespuestas.count()}')

        print(f'{c}/{eProfesoresMateria.count()}')
        c += 1
except Exception as ex:
    print(ex)
    pass


# try:
#     leadsselect = [73001, 73749, 75710, 73525, 72495, 72680, 75933, 74399, 72587, 75707, 75923, 75383, 73673, 73705, 73977, 73676, 71379, 73808, 74238, 73634, 75922, 72813, 73793, 73843, 72951, 73512, 72545, 76069, 73330, 73964, 72567, 74054, 74061, 73802, 75051, 73884, 70563, 73798, 48911, 73134, 72772, 70560, 73801, 75187, 73754, 73963, 73603, 75974, 74048, 74252, 76535, 73767, 73882, 72547, 72392, 73677, 75771]
#     cohorte  = None
#     c = 0
#     for le in leadsselect:
#         lead = InscripcionCohorte.objects.get(pk=le)
#
#         listarequisitos = EvidenciaRequisitosAspirante.objects.filter(inscripcioncohorte=lead, requisitos__requisito__claserequisito__clasificacion=1)
#         if lead.cohortes.maestriaadmision.carrera.id == 215:
#             cohorte = CohorteMaestria.objects.get(id=194)
#         elif lead.cohortes.maestriaadmision.carrera.id == 266:
#             cohorte = CohorteMaestria.objects.get(id=205)
#
#         canrequi = RequisitosMaestria.objects.filter(status=True, obligatorio=True,cohorte=cohorte,
#                                                      requisito__claserequisito__clasificacion=1).distinct().count()
#
#         listarequisitosfinan = EvidenciaRequisitosAspirante.objects.filter(inscripcioncohorte=lead, requisitos__requisito__claserequisito__clasificacion=3)
#
#         if lead.subirrequisitogarante:
#             canrequifi = RequisitosMaestria.objects.filter(status=True, obligatorio=True,
#                                                            cohorte=cohorte,
#                                                            requisito__claserequisito__clasificacion=3).distinct().count()
#         else:
#             canrequifi = RequisitosMaestria.objects.filter(status=True, obligatorio=True,
#                                                            cohorte=cohorte,
#                                                            requisito__claserequisito__clasificacion=3).exclude(requisito__id__in=[56, 57, 59]).distinct().count()
#
#         # SI TIENE EVIDENCIAS DE ADMISION EN LA COHORTE PASADA
#         for lis in listarequisitos:
#             if RequisitosMaestria.objects.filter(requisito=lis.requisitos.requisito,
#                                                  cohorte=cohorte, status=True):
#                 requi = RequisitosMaestria.objects.filter(requisito=lis.requisitos.requisito, cohorte=cohorte, status=True)[0]
#
#                 lis.requisitos = requi
#                 # if requi.requisito.id in (2, 14, 16, 4, 29, 52, 6, 62, 54):
#                 #     if lis.ultima_evidencia():
#                 #         dera = DetalleEvidenciaRequisitosAspirante.objects.get(pk=lis.ultima_evidencia().id)
#                 #         dera.estado_aprobacion = 1
#                 #         dera.observacion = "Pendiente debido a cambio de Cohorte"
#                 #         dera.save()
#                 lis.save()
#
#         canevi = EvidenciaRequisitosAspirante.objects.filter(inscripcioncohorte=lead,
#                                                              requisitos__cohorte=cohorte,
#                                                              requisitos__requisito__claserequisito__clasificacion=1,
#                                                              detalleevidenciarequisitosaspirante__estado_aprobacion=2,
#                                                              requisitos__status=True).distinct().count()
#
#         if canevi == canrequi:
#             lead.estado_aprobador = 2
#         else:
#             lead.estado_aprobador = 1
#             lead.todosubido = False
#             lead.preaprobado = False
#
#         lead.save()
#
#         # SI TIENE EVIDENCIAS DE FINANCIAMIENTO EN LA COHORTE PASADA
#         if lead.formapagopac:
#             if lead.formapagopac.id == 2:
#                 for listf in listarequisitosfinan:
#                     if RequisitosMaestria.objects.filter(requisito=listf.requisitos.requisito,
#                                                          cohorte=cohorte, status=True,
#                                                          requisito__claserequisito__clasificacion=3):
#                         requifinan = RequisitosMaestria.objects.filter(requisito=listf.requisitos.requisito,
#                                                                        cohorte=cohorte,
#                                                                        status=True,
#                                                                        requisito__claserequisito__clasificacion=3)[0]
#                         listf.requisitos = requifinan
#                         # if requifinan.requisito.id in (2, 14, 16, 4, 29, 52, 6, 62, 54):
#                         #     if listf.ultima_evidencia():
#                         #         dera = DetalleEvidenciaRequisitosAspirante.objects.get(
#                         #             pk=listf.ultima_evidencia().id)
#                         #         dera.estado_aprobacion = 1
#                         #         dera.observacion = "Pendiente debido a cambio de Cohorte"
#                         #         dera.save()
#                         listf.save()
#
#                 canevifi = EvidenciaRequisitosAspirante.objects.filter(inscripcioncohorte=lead,
#                                                                        requisitos__cohorte=cohorte,
#                                                                        requisitos__requisito__claserequisito__clasificacion=3,
#                                                                        detalleevidenciarequisitosaspirante__estado_aprobacion=2).count()
#
#                 if canevifi == canrequifi:
#                     lead.estadoformapago = 2
#                 else:
#                     lead.estadoformapago = 1
#                     lead.todosubidofi = False
#
#                 lead.save()
#
#         # SI TIENE RUBROS GENERADOS
#         chorte = CohorteMaestria.objects.get(id=cohorte.id, status=True)
#         if Rubro.objects.filter(inscripcion=lead, status=True).exists():
#             rubros = Rubro.objects.filter(inscripcion=lead, status=True)
#             tiporubroarancel = TipoOtroRubro.objects.get(pk=chorte.tiporubro.id)
#             for rubro in rubros:
#                 rubro.nombre = tiporubroarancel.nombre + ' - ' + chorte.maestriaadmision.descripcion + ' - ' + chorte.descripcion
#                 rubro.cohortemaestria = chorte
#                 rubro.save()
#
#                 if rubro.idrubroepunemi != 0:
#                     cursor = connections['epunemi'].cursor()
#                     sql = "UPDATE sagest_rubro SET nombre = '%s' WHERE sagest_rubro.status= true and sagest_rubro.id= %s" % (rubro.nombre, rubro.idrubroepunemi)
#                     cursor.execute(sql)
#                     cursor.close()
#
#         if Contrato.objects.filter(status=True, inscripcion=lead).exists():
#             contratopos = Contrato.objects.get(status=True, inscripcion=lead)
#
#             detalleevidencia = DetalleAprobacionContrato(contrato_id=contratopos.id, espagare=False,
#                                                          observacion='Rechazado por cambio de cohorte',
#                                                          persona_id=1, estado_aprobacion=3,
#                                                          fecha_aprobacion=datetime.now(),
#                                                          archivocontrato=contratopos.archivocontrato)
#             detalleevidencia.save()
#
#             if contratopos.inscripcion.formapagopac.id == 2:
#                 detalleevidencia = DetalleAprobacionContrato(contrato_id=contratopos.id, espagare=True,
#                                                              observacion='Rechazado por cambio de cohorte',
#                                                              persona_id=1, estado_aprobacion=3,
#                                                              fecha_aprobacion=datetime.now(),
#                                                              archivocontrato=contratopos.archivopagare)
#                 detalleevidencia.save()
#
#             contratopos.estado = 3
#             contratopos.estadopagare = 3
#             contratopos.save()
#
#         observacion = f'Cambio de {lead.cohortes} a {chorte}.'
#         cambio = CambioAdmitidoCohorteInscripcion(inscripcionCohorte=lead, cohortes=chorte,
#                                                   observacion=observacion)
#         cambio.save()
#
#         lead.cohortes_id = chorte.id
#         lead.save()
#         c +=1
#         print(f"{c} - {lead.inscripcionaspirante.persona.cedula} - {lead.__str__()} - {lead.cohortes.__str__()} - {lead.estado_aprobador}")
# except Exception as ex:
#     pass

# def imprimirreporte():
#     try:
#         __author__ = 'Unemi'
#
#         output = io.BytesIO()
#         name_document = 'reporte_plantilla'
#         nombre_archivo = name_document + "_11.xlsx"
#         directory = os.path.join(MEDIA_ROOT, 'reportes', 'gestion', nombre_archivo)
#         workbook = xlsxwriter.Workbook(directory, {'constant_memory': True})
#         ws = workbook.add_worksheet('plantilla')
#         ws.set_column(0, 0, 10)
#         ws.set_column(1, 1, 15)
#         ws.set_column(2, 2, 25)
#         ws.set_column(3, 3, 25)
#         ws.set_column(4, 4, 15)
#         ws.set_column(5, 5, 25)
#         ws.set_column(6, 6, 25)
#         ws.set_column(7, 7, 15)
#         ws.set_column(8, 8, 15)
#
#         formatotitulo_filtros = workbook.add_format(
#             {'bold': 1, 'text_wrap': True, 'border': 1, 'align': 'center', 'font_size': 14})
#
#         formatoceldacab = workbook.add_format(
#             {'align': 'center', 'bold': 1, 'border': 1, 'text_wrap': True, 'fg_color': '#1C3247',
#              'font_color': 'white'})
#         formatoceldaleft = workbook.add_format(
#             {'text_wrap': True, 'align': 'center', 'valign': 'vcenter', 'border': 1})
#
#         formatoceldaleft2 = workbook.add_format(
#             {'num_format': '$ #,##0.00', 'text_wrap': True, 'align': 'center', 'valign': 'vcenter', 'border': 1,
#              'bold': 1})
#
#         formatoceldaleft3 = workbook.add_format(
#             {'text_wrap': True, 'align': 'right', 'valign': 'vcenter', 'border': 1, 'bold': 1})
#
#         decimalformat = workbook.add_format(
#             {'num_format': '$ #,##0.00', 'text_wrap': True, 'align': 'center', 'valign': 'vcenter', 'border': 1})
#
#         decimalformat2 = workbook.add_format(
#             {'num_format': '$ #,##0.00', 'text_wrap': True, 'align': 'center', 'valign': 'vcenter', 'border': 1,
#              'bold': 1})
#
#         ws.write(2, 0, 'num', formatoceldacab)
#         ws.write(2, 1, 'identification', formatoceldacab)
#         ws.write(2, 2, 'first_name', formatoceldacab)
#         ws.write(2, 3, 'last_name', formatoceldacab)
#         ws.write(2, 4, 'alias', formatoceldacab)
#         ws.write(2, 5, 'role', formatoceldacab)
#         ws.write(2, 6, 'gender', formatoceldacab)
#         ws.write(2, 7, 'birth_date', formatoceldacab)
#         ws.write(2, 8, '¿tiene foto?', formatoceldacab)
#
#         filtro = Q(status=True)
#
#         plantillas = DistributivoPersona.objects.filter(filtro).order_by('id')
#
#         filas_recorridas = 4
#         c = 0
#         for plantilla in plantillas:
#             rol = ''
#             if plantilla.denominacionpuesto.id in [580, 61, 561, 50, 736, 564, 966, 819, 821, 34, 553, 973, 40, 722, 47,
#                                                    735, 568, 741, 681, 540, 678, 463, 952, 954, 45, 578, 720, 945, 827,
#                                                    550, 559, 739, 721, 51, 967, 832, 554, 788, 402, 378, 278, 959, 794,
#                                                    752, 962, 825, 814, 732, 900, 839, 946, 841, 792, 724, 969, 757, 793,
#                                                    731, 755, 686, 971, 751, 16, 958, 772, 577, 530, 899, 733, 661, 659,
#                                                    669, 964, 759, 754, 750, 833, 532, 30, 471, 807, 517, 956, 710, 526,
#                                                    762, 790, 761, 789, 824, 970, 691, 791, 636, 955, 826, 383, 901, 902,
#                                                    965, 766, 76, 616, 713, 515, 143, 548, 740, 823, 822, 968, 169, 398,
#                                                    397, 118, 120, 717, 799, 947, 648, 646, 137, 150, 805, 144, 132, 829,
#                                                    838, 149, 645, 642, 649, 644, 803, 782, 650, 647, 695, 138, 699, 831,
#                                                    702, 703, 804, 781, 944, 801, 800, 817, 816, 809, 780, 779, 492, 778,
#                                                    719, 815, 811, 728, 718, 598, 716, 612, 611, 376, 508, 510, 776, 87,
#                                                    469, 83, 599, 602, 709, 586, 711, 786, 961, 569, 725, 963, 960, 768,
#                                                    573, 747, 769, 810, 743, 734, 957, 953, 565, 787, 773, 742, 840, 567,
#                                                    834, 802, 72, 66, 478, 176, 535, 74, 113, 502, 972, 796, 795, 797]:
#                 rol = 'ADMINISTRATIVO'
#             elif plantilla.denominacionpuesto.id in [101, 99, 360, 812, 359, 715, 104, 95, 400, 504, 102, 97, 190, 189,
#                                                      362, 399, 466]:
#                 rol = 'DOCENTE'
#             elif plantilla.denominacionpuesto.id in [622, 836, 623, 109, 533, 107, 619, 4, 2, 3, 111, 639, 584, 828,
#                                                      582, 108, 730, 536, 624, 837, 693, 506, 638, 637, 640, 621, 618,
#                                                      78, 188, 365, 547, 556, 641]:
#                 rol = 'TRABAJADOR'
#
#             sexo = ''
#             if plantilla.persona.sexo.id == 1:
#                 sexo = 'MUJER'
#             elif plantilla.persona.sexo.id == 2:
#                 sexo = 'HOMBRE'
#
#             ws.write('A%s' % filas_recorridas, str(plantilla.id), formatoceldaleft)
#             ws.write('B%s' % filas_recorridas, str(plantilla.persona.cedula), formatoceldaleft)
#             ws.write('C%s' % filas_recorridas, str(plantilla.persona.nombres), formatoceldaleft)
#             ws.write('D%s' % filas_recorridas, str(plantilla.persona.apellido1 + ' ' + plantilla.persona.apellido2), formatoceldaleft)
#             ws.write('E%s' % filas_recorridas, str(''), formatoceldaleft)
#             ws.write('F%s' % filas_recorridas, str(rol), formatoceldaleft)
#             ws.write('G%s' % filas_recorridas, str(sexo), formatoceldaleft)
#             ws.write('H%s' % filas_recorridas, str(plantilla.persona.nacimiento.strftime("%d/%m/%Y")), formatoceldaleft)
#             ws.write('I%s' % filas_recorridas, str('SI' if plantilla.persona.foto() else 'NO'), formatoceldaleft)
#
#             filas_recorridas += 1
#             c += 1
#             print(f'{c}/{plantillas.count()} - {plantilla.persona}')
#
#         workbook.close()
#         # output.seek(0)
#         # fecha_hora_actual = datetime.now().date()
#         # filename = 'Listado_ventas_' + str(fecha_hora_actual) + '.xlsx'
#         response = HttpResponse(directory, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#         response['Content-Disposition'] = 'attachment; filename=%s' % name_document
#         url_file = "{}reportes/gestion/{}".format(MEDIA_URL, nombre_archivo)
#         print(url_file)
#     except Exception as ex:
#         pass
#
# imprimirreporte()

# print(u"********************COHORTES BACHILLERATO CIENCIAS NATURALES***************************")
# try:
#     cohortes = CohorteMaestria.objects.filter(status=True, maestriaadmision__carrera__id=236).exclude(pk=192)
#     maestriaadmision = MaestriasAdmision.objects.get(status=True, carrera__id=267)
#     valorexamen = 0
#     for cohorte in cohortes:
#         print(f'Cohorte a duplicar: {cohorte}')
#         periodo = None
#         if cohorte.id == 162:
#             if Periodo.objects.filter(nombre='MAESTRÍA EN EDUCACIÓN DE BACHILLERATO CON MENCIÓN EN PEDAGOGÍA DE LAS CIENCIAS NATURALES - COHORTE I 2023').exists():
#                 periodo = Periodo.objects.get(nombre='MAESTRÍA EN EDUCACIÓN DE BACHILLERATO CON MENCIÓN EN PEDAGOGÍA DE LAS CIENCIAS NATURALES - COHORTE I 2023')
#         elif cohorte.id == 179:
#             if Periodo.objects.filter(nombre='MAESTRÍA EN EDUCACIÓN DE BACHILLERATO CON MENCIÓN EN PEDAGOGÍA DE LAS CIENCIAS NATURALES - COHORTE II 2023').exists():
#                 periodo = Periodo.objects.get(nombre='MAESTRÍA EN EDUCACIÓN DE BACHILLERATO CON MENCIÓN EN PEDAGOGÍA DE LAS CIENCIAS NATURALES - COHORTE II 2023')
#
#         if periodo:
#             if not CohorteMaestria.objects.filter(status=True, periodoacademico=periodo).exists():
#                 programamaestria = CohorteMaestria(periodoacademico=periodo,
#                                                     coordinador=cohorte.coordinador,
#                                                     maestriaadmision=maestriaadmision,
#                                                     descripcion=cohorte.descripcion,
#                                                     modalidad=cohorte.modalidad,
#                                                     alias=cohorte.alias,
#                                                     numerochorte=cohorte.numerochorte,
#                                                     cupodisponible=cohorte.cupodisponible,
#                                                     cantidadgruposexamen=cohorte.cantidadgruposexamen,
#                                                     fechainiciocohorte=cohorte.fechainiciocohorte,
#                                                     fechafincohorte=cohorte.fechafincohorte,
#                                                     fechainicioinsp=cohorte.fechainicioinsp,
#                                                     fechafininsp=cohorte.fechafininsp,
#                                                     fechainiciorequisito=cohorte.fechainiciorequisito,
#                                                     fechafinrequisito=cohorte.fechafinrequisito,
#                                                     fechafinrequisitobeca=cohorte.fechafinrequisitobeca,
#                                                     fechainicioexamen=cohorte.fechainicioexamen,
#                                                     fechafinexamen=cohorte.fechafinexamen,
#                                                     notaminimaexa=cohorte.notaminimaexa,
#                                                     notaminimatest=cohorte.notaminimatest,
#                                                     valorexamen=valorexamen,
#                                                     valortramite=cohorte.valortramite,
#                                                     activo=cohorte.activo,
#                                                     minutosrango=cohorte.minutosrango,
#                                                     cantidadgruposentrevista=cohorte.cantidadgruposentrevista,
#                                                     totaladmitidoscohorte=cohorte.totaladmitidoscohorte)
#                 programamaestria.save()
#
#                 print(f'Nueva cohorte creada: {programamaestria}')
#
#                 if cohorte.tienecostomatricula:
#                     programamaestria.tienecostomatricula = cohorte.tienecostomatricula
#                     programamaestria.valormatricula = cohorte.valormatricula
#                 else:
#                     programamaestria.tienecostomatricula = False
#                     programamaestria.valormatricula = 0
#                 if cohorte.tienecostototal:
#                     programamaestria.tienecostototal = cohorte.tienecostototal
#                     programamaestria.valorprograma = cohorte.valorprograma
#                     # if f.cleaned_data['tipootrorubro']:
#                     #     cohorte.tiporubro_id = f.cleaned_data['tipootrorubro']
#                 else:
#                     programamaestria.tienecostototal = False
#                     programamaestria.valorprograma = 0
#                     programamaestria.tiporubro_id = None
#                 if cohorte.tiporubro:
#                     programamaestria.tiporubro_id = cohorte.tiporubro.id
#
#                 programamaestria.valorprogramacertificado = cohorte.valorprogramacertificado
#                 programamaestria.fechavencerubro = cohorte.fechavencerubro
#                 programamaestria.fechainiordinaria = cohorte.fechainiordinaria
#                 programamaestria.fechafinordinaria = cohorte.fechafinordinaria
#                 programamaestria.fechainiextraordinaria = cohorte.fechainiextraordinaria
#                 programamaestria.fechafinextraordinaria = cohorte.fechafinextraordinaria
#                 programamaestria.presupuestobeca = cohorte.presupuestobeca
#                 programamaestria.save()
#
#                 print(f'Rubro configurado')
#
#                 requisitos = RequisitosMaestria.objects.filter(status=True, cohorte=cohorte)
#                 for requisitom in requisitos:
#                     if not RequisitosMaestria.objects.filter(cohorte=programamaestria, requisito=requisitom.requisito,
#                                                              status=True).exists():
#                         requisitomaestria = RequisitosMaestria(cohorte=programamaestria,
#                                                                requisito=requisitom.requisito)
#                         requisitomaestria.save()
#
#                 print(f'Requisitos configurado')
#
#                 tiposfinancia = ConfigFinanciamientoCohorte.objects.filter(status=True, cohorte=cohorte)
#                 for tipo in tiposfinancia:
#                     instance = ConfigFinanciamientoCohorte(descripcion=tipo.descripcion,
#                                                             cohorte_id=programamaestria.id,
#                                                             valormatricula=tipo.valormatricula,
#                                                             valorarancel=tipo.valorarancel,
#                                                             valortotalprograma=tipo.valortotalprograma,
#                                                             porcentajeminpagomatricula=tipo.porcentajeminpagomatricula,
#                                                             maxnumcuota=tipo.maxnumcuota,
#                                                             fecha=tipo.fecha)
#                     instance.save()
#
#                 print(f'Tipos de financiamiento creados')
#
#                 asesores = AsesorMeta.objects.filter(status=True, cohorte=cohorte)
#
#                 for asesorm in asesores:
#                     if not AsesorMeta.objects.filter(status=True, cohorte=cohorte, asesor=asesorm.asesor).exists():
#                         asesormeta = AsesorMeta(asesor=asesorm.asesor,
#                                                 cohorte=programamaestria,
#                                                 fecha_inicio_meta=programamaestria.fechainicioinsp,
#                                                 fecha_fin_meta=programamaestria.fechafininsp,
#                                                 meta=0)
#                         asesormeta.save()
#
#                 print(f'Asesor Meta creado')
#
#     print(u"********************COHORTES BACHILLERATO CIENCIAS SOCIALES***************************")
#     cohortes = CohorteMaestria.objects.filter(status=True, maestriaadmision__carrera__id=236).exclude(pk=192)
#     maestriaadmision = MaestriasAdmision.objects.get(status=True, carrera__id=268)
#     valorexamen = 0
#     for cohorte in cohortes:
#         print(f'Cohorte a duplicar: {cohorte}')
#         periodo = None
#         if cohorte.id == 162:
#             if Periodo.objects.filter(nombre='MAESTRÍA EN EDUCACIÓN DE BACHILLERATO MENCIÓN EN PEDAGOGÍA EN CIENCIAS SOCIALES - COHORTE I 2023').exists():
#                 periodo = Periodo.objects.get(nombre='MAESTRÍA EN EDUCACIÓN DE BACHILLERATO MENCIÓN EN PEDAGOGÍA EN CIENCIAS SOCIALES - COHORTE I 2023')
#         elif cohorte.id == 179:
#             if Periodo.objects.filter(nombre='MAESTRÍA EN EDUCACIÓN DE BACHILLERATO MENCIÓN EN PEDAGOGÍA EN CIENCIAS SOCIALES - COHORTE II 2023').exists():
#                 periodo = Periodo.objects.get(nombre='MAESTRÍA EN EDUCACIÓN DE BACHILLERATO MENCIÓN EN PEDAGOGÍA EN CIENCIAS SOCIALES - COHORTE II 2023')
#
#         if periodo:
#             if not CohorteMaestria.objects.filter(status=True, periodoacademico=periodo).exists():
#                 programamaestria = CohorteMaestria(periodoacademico=periodo,
#                                                     coordinador=cohorte.coordinador,
#                                                     maestriaadmision=maestriaadmision,
#                                                     descripcion=cohorte.descripcion,
#                                                     modalidad=cohorte.modalidad,
#                                                     alias=cohorte.alias,
#                                                     numerochorte=cohorte.numerochorte,
#                                                     cupodisponible=cohorte.cupodisponible,
#                                                     cantidadgruposexamen=cohorte.cantidadgruposexamen,
#                                                     fechainiciocohorte=cohorte.fechainiciocohorte,
#                                                     fechafincohorte=cohorte.fechafincohorte,
#                                                     fechainicioinsp=cohorte.fechainicioinsp,
#                                                     fechafininsp=cohorte.fechafininsp,
#                                                     fechainiciorequisito=cohorte.fechainiciorequisito,
#                                                     fechafinrequisito=cohorte.fechafinrequisito,
#                                                     fechafinrequisitobeca=cohorte.fechafinrequisitobeca,
#                                                     fechainicioexamen=cohorte.fechainicioexamen,
#                                                     fechafinexamen=cohorte.fechafinexamen,
#                                                     notaminimaexa=cohorte.notaminimaexa,
#                                                     notaminimatest=cohorte.notaminimatest,
#                                                     valorexamen=valorexamen,
#                                                     valortramite=cohorte.valortramite,
#                                                     activo=cohorte.activo,
#                                                     minutosrango=cohorte.minutosrango,
#                                                     cantidadgruposentrevista=cohorte.cantidadgruposentrevista,
#                                                     totaladmitidoscohorte=cohorte.totaladmitidoscohorte)
#                 programamaestria.save()
#
#                 print(f'Nueva cohorte creada: {programamaestria}')
#
#                 if cohorte.tienecostomatricula:
#                     programamaestria.tienecostomatricula = cohorte.tienecostomatricula
#                     programamaestria.valormatricula = cohorte.valormatricula
#                 else:
#                     programamaestria.tienecostomatricula = False
#                     programamaestria.valormatricula = 0
#                 if cohorte.tienecostototal:
#                     programamaestria.tienecostototal = cohorte.tienecostototal
#                     programamaestria.valorprograma = cohorte.valorprograma
#                     # if f.cleaned_data['tipootrorubro']:
#                     #     cohorte.tiporubro_id = f.cleaned_data['tipootrorubro']
#                 else:
#                     programamaestria.tienecostototal = False
#                     programamaestria.valorprograma = 0
#                     programamaestria.tiporubro_id = None
#                 if cohorte.tiporubro:
#                     programamaestria.tiporubro_id = cohorte.tiporubro.id
#
#                 programamaestria.valorprogramacertificado = cohorte.valorprogramacertificado
#                 programamaestria.fechavencerubro = cohorte.fechavencerubro
#                 programamaestria.fechainiordinaria = cohorte.fechainiordinaria
#                 programamaestria.fechafinordinaria = cohorte.fechafinordinaria
#                 programamaestria.fechainiextraordinaria = cohorte.fechainiextraordinaria
#                 programamaestria.fechafinextraordinaria = cohorte.fechafinextraordinaria
#                 programamaestria.presupuestobeca = cohorte.presupuestobeca
#                 programamaestria.save()
#
#                 print(f'Rubro configurado')
#
#                 requisitos = RequisitosMaestria.objects.filter(status=True, cohorte=cohorte)
#                 for requisitom in requisitos:
#                     if not RequisitosMaestria.objects.filter(cohorte=programamaestria, requisito=requisitom.requisito,
#                                                              status=True).exists():
#                         requisitomaestria = RequisitosMaestria(cohorte=programamaestria,
#                                                                requisito=requisitom.requisito)
#                         requisitomaestria.save()
#
#                 print(f'Requisitos configurado')
#
#                 tiposfinancia = ConfigFinanciamientoCohorte.objects.filter(status=True, cohorte=cohorte)
#                 for tipo in tiposfinancia:
#                     instance = ConfigFinanciamientoCohorte(descripcion=tipo.descripcion,
#                                                             cohorte_id=programamaestria.id,
#                                                             valormatricula=tipo.valormatricula,
#                                                             valorarancel=tipo.valorarancel,
#                                                             valortotalprograma=tipo.valortotalprograma,
#                                                             porcentajeminpagomatricula=tipo.porcentajeminpagomatricula,
#                                                             maxnumcuota=tipo.maxnumcuota,
#                                                             fecha=tipo.fecha)
#                     instance.save()
#
#                 print(f'Tipos de financiamiento creados')
#
#                 asesores = AsesorMeta.objects.filter(status=True, cohorte=cohorte)
#
#                 for asesorm in asesores:
#                     if not AsesorMeta.objects.filter(status=True, cohorte=cohorte, asesor=asesorm.asesor).exists():
#                         asesormeta = AsesorMeta(asesor=asesorm.asesor,
#                                                 cohorte=programamaestria,
#                                                 fecha_inicio_meta=programamaestria.fechainicioinsp,
#                                                 fecha_fin_meta=programamaestria.fechafininsp,
#                                                 meta=0)
#                         asesormeta.save()
#
#                 print(f'Asesor Meta creado')
# except Exception as ex:
#     pass

# print(u"UNIDADES CERTIFICADORAS")
# try:
#     certificados = Certificado.objects.filter(status=True, visible=True, servicio__isnull=False).exclude(tipo_origen=1)
#     responsable = Persona.objects.get(pk=127492)
#     for certificado in certificados:
#         if CertificadoUnidadCertificadora.objects.filter(certificado=certificado, responsable__id=29110).exists():
#             unidad = CertificadoUnidadCertificadora.objects.get(certificado=certificado, responsable__id=29110)
#             unidad.responsable = responsable
#             unidad.responsable_titulo = 'Abg. Edison Sempertegui Henriquez'
#             unidad.responsable_denominacion = 'Secretario General (S)'
#             unidad.save()
#             print(f'{certificado} - {certificado.codigo} - Responsable: {unidad.responsable} - Titulo: {unidad.responsable_titulo} - Cargo: {unidad.responsable_denominacion}')
# except Exception as ex:
#     pass
# try:
#     cohortes = CohorteMaestria.objects.filter(status=True, procesoabierto=False, maestriaadmision__carrera__id=215).exclude(pk=194)
#     for cohorte in cohortes:
#         if CohorteMaestria.objects.filter(status=True, maestriaadmision__carrera__id=266, descripcion=cohorte.descripcion, procesoabierto=False).exists():
#             cohortenueva = CohorteMaestria.objects.get(status=True, maestriaadmision__carrera__id=266, descripcion=cohorte.descripcion, procesoabierto=False)
#             postulantes = InscripcionCohorte.objects.filter(status=True, cohortes=cohorte, id__in=[]).exclude(cohortes__id__in=[194, 205]).order_by('id')
#             c = 0
#             for postulante in postulantes:
#                 postulante.cohortes = cohortenueva
#                 postulante.save()
#
#                 requisitos = RequisitosMaestria.objects.filter(status=True, cohorte=cohortenueva)
#
#                 for reqma in requisitos:
#                     if EvidenciaRequisitosAspirante.objects.filter(inscripcioncohorte=postulante, requisitos__requisito=reqma.requisito).exists():
#                         evi = EvidenciaRequisitosAspirante.objects.filter(inscripcioncohorte=postulante, requisitos__requisito=reqma.requisito).first()
#                         evi.requisitos = reqma
#                         evi.save()
#
#                 if Rubro.objects.filter(status=True, inscripcion=postulante).exists():
#                     rubros = Rubro.objects.filter(status=True, inscripcion=postulante)
#                     for rubro in rubros:
#                         rubro.cohortemaestria = cohortenueva
#                         rubro.tipo = cohorte.tiporubro
#                         rubro.save()
#                 if postulante.Configfinanciamientocohorte:
#                     if ConfigFinanciamientoCohorte.objects.filter(descripcion=postulante.Configfinanciamientocohorte.descripcion, cohorte=cohortenueva).exists():
#                         finan = ConfigFinanciamientoCohorte.objects.filter(descripcion=postulante.Configfinanciamientocohorte.descripcion, cohorte=cohortenueva).first()
#                         postulante.Configfinanciamientocohorte = finan
#                         postulante.save()
#                 c += 1
#                 print(f'{c}/{postulantes.count()} - Cedula: {postulante.inscripcionaspirante.persona.cedula} - {postulante}')
# except Exception as ex:
#     pass
# try:
#     c = 0
#     cohortenueva = CohorteMaestria.objects.get(status=True, pk=206)
#     postulantes = InscripcionCohorte.objects.filter(status=True, cohortes__id=192, itinerario=1).order_by('id')
#     for postulante in postulantes:
#         postulante.cohortes = cohortenueva
#         postulante.save()
#
#         requisitos = RequisitosMaestria.objects.filter(status=True, cohorte=cohortenueva)
#
#         for reqma in requisitos:
#             if EvidenciaRequisitosAspirante.objects.filter(inscripcioncohorte=postulante, requisitos__requisito=reqma.requisito).exists():
#                 evi = EvidenciaRequisitosAspirante.objects.filter(inscripcioncohorte=postulante, requisitos__requisito=reqma.requisito).first()
#                 evi.requisitos = reqma
#                 evi.save()
#
#         if Rubro.objects.filter(status=True, inscripcion=postulante).exists():
#             rubros = Rubro.objects.filter(status=True, inscripcion=postulante)
#             tipo = TipoOtroRubro.objects.get(pk=3566)
#             for rubro in rubros:
#                 rubro.cohortemaestria = cohortenueva
#                 rubro.tipo = tipo
#                 rubro.save()
#         if postulante.Configfinanciamientocohorte:
#             if ConfigFinanciamientoCohorte.objects.filter(descripcion__icontains=postulante.Configfinanciamientocohorte.descripcion, cohorte=cohortenueva).exists():
#                 finan = ConfigFinanciamientoCohorte.objects.filter(descripcion__icontains=postulante.Configfinanciamientocohorte.descripcion, cohorte=cohortenueva).first()
#                 postulante.Configfinanciamientocohorte = finan
#                 postulante.save()
#         c += 1
#         print(f'{c}/{postulantes.count()} - Cedula: {postulante.inscripcionaspirante.persona.cedula} - {postulante}')
#
#     print('*************************+++CIENCIAS SOCIALES - ACTIVOS*************************')
#     c = 0
#     cohortenueva = CohorteMaestria.objects.get(status=True, pk=207)
#     postulantes = InscripcionCohorte.objects.filter(status=True, cohortes__id=192, itinerario=3).order_by('id')
#     for postulante in postulantes:
#         postulante.cohortes = cohortenueva
#         postulante.save()
#
#         requisitos = RequisitosMaestria.objects.filter(status=True, cohorte=cohortenueva)
#
#         for reqma in requisitos:
#             if EvidenciaRequisitosAspirante.objects.filter(inscripcioncohorte=postulante, requisitos__requisito=reqma.requisito).exists():
#                 evi = EvidenciaRequisitosAspirante.objects.filter(inscripcioncohorte=postulante, requisitos__requisito=reqma.requisito).first()
#                 evi.requisitos = reqma
#                 evi.save()
#
#         if Rubro.objects.filter(status=True, inscripcion=postulante).exists():
#             rubros = Rubro.objects.filter(status=True, inscripcion=postulante)
#             tipo = TipoOtroRubro.objects.get(pk=3565)
#             for rubro in rubros:
#                 rubro.cohortemaestria = cohortenueva
#                 rubro.tipo = tipo
#                 rubro.save()
#         if postulante.Configfinanciamientocohorte:
#             if ConfigFinanciamientoCohorte.objects.filter(descripcion__icontains=postulante.Configfinanciamientocohorte.descripcion, cohorte=cohortenueva).exists():
#                 finan = ConfigFinanciamientoCohorte.objects.filter(descripcion__icontains=postulante.Configfinanciamientocohorte.descripcion, cohorte=cohortenueva).first()
#                 postulante.Configfinanciamientocohorte = finan
#                 postulante.save()
#         c += 1
#         print(f'{c}/{postulantes.count()} - Cedula: {postulante.inscripcionaspirante.persona.cedula} - {postulante}')
# except Exception as ex:
#     pass
# def promedio_re(inscripcion):
#     return round(null_to_numeric(
#         inscripcion.recordacademico_set.filter(validapromedio=True, aprobada=True, status=True, asignaturamalla__isnull=False).exclude(asignaturamalla__nivelmalla=inscripcion.mi_nivel().nivel).aggregate(
#             promedio=Avg('nota'))['promedio']), 2)
#
#
# def imprimirreporte():
#     try:
#         estado = ''
#
#         __author__ = 'Unemi'
#
#         output = io.BytesIO()
#         name_document = 'reporte_mejores_estudiantes_facultad'
#         nombre_archivo = name_document + "_11.xlsx"
#         directory = os.path.join(MEDIA_ROOT, 'reportes', 'gestion', nombre_archivo)
#         workbook = xlsxwriter.Workbook(directory, {'constant_memory': True})
#         ws = workbook.add_worksheet('resultados')
#         ws.set_column(0, 0, 10)
#         ws.set_column(1, 1, 15)
#         ws.set_column(2, 2, 35)
#         ws.set_column(3, 3, 45)
#         ws.set_column(4, 4, 25)
#         ws.set_column(5, 5, 15)
#         ws.set_column(6, 6, 45)
#         ws.set_column(7, 7, 15)
#
#         formatotitulo_filtros = workbook.add_format(
#             {'bold': 1, 'text_wrap': True, 'border': 1, 'align': 'center', 'font_size': 14})
#
#         formatoceldacab = workbook.add_format(
#             {'align': 'center', 'bold': 1, 'border': 1, 'text_wrap': True, 'fg_color': '#1C3247', 'font_color': 'white'})
#         formatoceldaleft = workbook.add_format(
#             {'text_wrap': True, 'align': 'center', 'valign': 'vcenter', 'border': 1})
#
#         formatoceldaleft2 = workbook.add_format(
#             {'num_format': '$ #,##0.00', 'text_wrap': True, 'align': 'center', 'valign': 'vcenter', 'border': 1, 'bold': 1})
#
#         formatoceldaleft3 = workbook.add_format(
#             {'text_wrap': True, 'align': 'right', 'valign': 'vcenter', 'border': 1, 'bold': 1})
#
#         decimalformat = workbook.add_format(
#             {'num_format': '#,##0.00', 'text_wrap': True, 'align': 'center', 'valign': 'vcenter', 'border': 1})
#
#         decimalformat2 = workbook.add_format(
#             {'num_format': '$ #,##0.00', 'text_wrap': True, 'align': 'center', 'valign': 'vcenter', 'border': 1, 'bold': 1})
#
#         ws.write(2, 0, 'N°', formatoceldacab)
#         ws.write(2, 1, 'Idmat', formatoceldacab)
#         ws.write(2, 2, 'Coordinacion', formatoceldacab)
#         ws.write(2, 3, 'Carrera', formatoceldacab)
#         ws.write(2, 4, 'Nivel', formatoceldacab)
#         ws.write(2, 5, 'Cédula', formatoceldacab)
#         ws.write(2, 6, 'Maestrante', formatoceldacab)
#         ws.write(2, 7, 'Promedio', formatoceldacab)
#
#         filas_recorridas = 4
#         cont = 1
#         conta = 1
#
#         diccionario_promedios = {}
#         matriculados = Matricula.objects.filter(status=True, inscripcion__status=True,
#                                                 inscripcion__carrera__status=True, retiradomatricula=False,
#                                                 nivel__periodo__id=224, nivelmalla__id__gte=4, inscripcion__carrera__coordinacion__id__in=[5]).distinct()
#         for matriculado in matriculados:
#             canasignaturasnivel = AsignaturaMalla.objects.filter(status=True, malla=matriculado.inscripcion.malla_inscripcion().malla,
#                                                          nivelmalla=matriculado.inscripcion.mi_nivel().nivel).distinct().count()
#
#             idasignaturasnivel = AsignaturaMalla.objects.filter(status=True, malla=matriculado.inscripcion.malla_inscripcion().malla,
#                                                          nivelmalla=matriculado.inscripcion.mi_nivel().nivel).values_list('id', flat=True)
#
#             canmateriasnivel = MateriaAsignada.objects.filter(status=True, matricula=matriculado,
#                                                                  materia__asignaturamalla__id__in=idasignaturasnivel).distinct().count()
#
#             if not HistoricoRecordAcademico.objects.filter(status=True, inscripcion=matriculado.inscripcion, aprobada=False).exists():
#                 if canasignaturasnivel == canmateriasnivel:
#                     nombre = matriculado.id
#                     numero = promedio_re(matriculado.inscripcion)
#                     diccionario_promedios[nombre] = numero
#                 else:
#                     print(f'{conta} / {matriculados.count()}')
#                     conta += 1
#             else:
#                 print(f'{conta} / {matriculados.count()}')
#                 conta += 1
#         diccionario_ordenado = dict(sorted(diccionario_promedios.items(), key=lambda item: item[1], reverse=True))
#         lista = []
#         for nombre in diccionario_ordenado.keys():
#             if len(lista) <= 5:
#                 lista.append(nombre)
#
#         for lis in lista:
#             matriculado = Matricula.objects.get(status=True, pk=lis)
#             ws.write('A%s' % filas_recorridas, str(cont), formatoceldaleft)
#             ws.write('B%s' % filas_recorridas, str(matriculado.id), formatoceldaleft)
#             ws.write('C%s' % filas_recorridas, str(matriculado.inscripcion.carrera.coordinacion_carrera().nombre), formatoceldaleft)
#             ws.write('D%s' % filas_recorridas, str(matriculado.inscripcion.carrera), formatoceldaleft)
#             ws.write('E%s' % filas_recorridas, str(matriculado.inscripcion.mi_nivel().nivel), formatoceldaleft)
#             ws.write('F%s' % filas_recorridas, str(matriculado.inscripcion.persona.cedula), formatoceldaleft)
#             ws.write('G%s' % filas_recorridas, str(matriculado.inscripcion.persona.nombre_completo_inverso()), formatoceldaleft)
#             ws.write('H%s' % filas_recorridas, promedio_re(matriculado.inscripcion), decimalformat)
#
#             filas_recorridas += 1
#             print(f'{cont} / {len(lista)}')
#             cont += 1
#
#         workbook.close()
#         # output.seek(0)
#         # fecha_hora_actual = datetime.now().date()
#         # filename = 'Listado_ventas_' + str(fecha_hora_actual) + '.xlsx'
#         response = HttpResponse(directory, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#         response['Content-Disposition'] = 'attachment; filename=%s' % name_document
#         url_file = "{}reportes/gestion/{}".format(MEDIA_URL, nombre_archivo)
#         print(url_file)
#     except Exception as ex:
#         pass
#
# imprimirreporte()

# def imprimirreporte():
#     try:
#         estado = ''
#
#         __author__ = 'Unemi'
#
#         output = io.BytesIO()
#         name_document = 'reporte_mariela'
#         nombre_archivo = name_document + "_8.xlsx"
#         directory = os.path.join(MEDIA_ROOT, 'reportes', 'gestion', nombre_archivo)
#         workbook = xlsxwriter.Workbook(directory, {'constant_memory': True})
#         ws = workbook.add_worksheet('PAUTAS MAESTRIAS')
#         ws.set_column(0, 0, 10)
#         ws.set_column(1, 1, 15)
#         ws.set_column(2, 2, 15)
#         ws.set_column(3, 3, 30)
#         ws.set_column(4, 4, 30)
#         ws.set_column(5, 5, 40)
#         ws.set_column(6, 6, 30)
#         ws.set_column(7, 7, 30)
#
#         formatotitulo_filtros = workbook.add_format(
#             {'bold': 1, 'text_wrap': True, 'border': 1, 'align': 'center', 'font_size': 14})
#
#         formatoceldacab = workbook.add_format(
#             {'align': 'center', 'bold': 1, 'border': 1, 'text_wrap': True, 'fg_color': '#1C3247', 'font_color': 'white'})
#         formatoceldaleft = workbook.add_format(
#             {'text_wrap': True, 'align': 'center', 'valign': 'vcenter', 'border': 1})
#
#         formatoceldaleft2 = workbook.add_format(
#             {'num_format': '$ #,##0.00', 'text_wrap': True, 'align': 'center', 'valign': 'vcenter', 'border': 1, 'bold': 1})
#
#         formatoceldaleft3 = workbook.add_format(
#             {'text_wrap': True, 'align': 'right', 'valign': 'vcenter', 'border': 1, 'bold': 1})
#
#         decimalformat = workbook.add_format(
#             {'num_format': '$ #,##0.00', 'text_wrap': True, 'align': 'center', 'valign': 'vcenter', 'border': 1})
#
#         decimalformat2 = workbook.add_format(
#             {'num_format': '$ #,##0.00', 'text_wrap': True, 'align': 'center', 'valign': 'vcenter', 'border': 1, 'bold': 1})
#
#         ws.write(2, 0, 'N°', formatoceldacab)
#         ws.write(2, 1, 'Id', formatoceldacab)
#         ws.write(2, 2, 'Cédula', formatoceldacab)
#         ws.write(2, 3, 'Postulante', formatoceldacab)
#         ws.write(2, 4, 'Maestría', formatoceldacab)
#         ws.write(2, 5, 'Cohorte', formatoceldacab)
#         ws.write(2, 6, 'Fecha de postulación', formatoceldacab)
#         ws.write(2, 7, 'Estado', formatoceldacab)
#         ws.write(2, 8, '¿Activo?', formatoceldacab)
#
#         postulantes = InscripcionCohorte.objects.filter(inscripcionaspirante__persona__cedula__in=['0926320755',
# '0804373322',
# '1724785488',
# '1207708395',
# '1250194055',
# '0202082509',
# '2300338940',
# '0705368751',
# '1600858292',
# '0955350418',
# '1317064119',
# '1724785488',
# '1725635591',
# '0701761595',
# '0705821254',
# '0106585540',
# '0924145592',
# '0930302625',
# '1718605809',
# '1707649438',
# '0503907420',
# '0705049013',
# '0605722594',
# '1724996234',
# '1600611675',
# '1313510818',
# '1725690604',
# '2100771233',
# '0929598464',
# '1314592161',
# '0704446012',
# '1150142808',
# '0803753201',
# '0932348964',
# '0105990840',
# '1206074583',
# '0201822707',
# '1207368380',
# '0921600326',
# '0605126853',
# '0927926386',
# '0705465201',
# '0926395039',
# '1724442650',
# '0705706919',
# '1726289679',
# '1314480235',
# '0930614714',
# '1501083040',
# '0953800505',
# '1501018020',
# '0943803585',
# '1105256463',
# '0929012565',
# '0107071151',
# '0953412236',
# '1313792531',
# '0923657407',
# '1206708099',
# '0850141094',
# '1709054819',
# '0952761559',
# '0930686969',
# '0202384475',
# '0927228643',
# '0954895595',
# '0923302897',
# '0941549164',
# '1205380379',
# '1721112140',
# '804148104',
# '1106011065',
# '1105811168',
# '803545896',
# '350133336',
# '1105117327',
# '1104979685',
# '105977862',
# '1715402499',
# '1313889022',
# '1717799801',
# '105224521',
# '1312631458',
# '1313206094',
# '302574678',
# '2200570501',
# '2450596966',
# '202193413',
# '106585540',
# '705860948',
# '931665186',
# '1805237946',
# '951218031',
# '931266605',
# '705860948',
# '1105811168',
# '1003361126',
# '1205427071',
# '1600677767',
# '927781153',
# '605821776',
# '923782882',
# '150674984',
# '1003505615',
# '350133336',
# '606132777',
# '1003642863',
# '1600526469',
# '201632536',
# '1718393059',
# '927228353',
# '604255307',
# '1723088983',
# '1804875035',
# '1105353377',
# '923351613',
# '1207628643',
# '929167674',
# '703879783',
# '705917201',
# '105977862',
# '106153216',
# '803818798',
# '803754829',
# '1716594856',
# '703994855',
# '1312611617',
# '705624765',
# '1900609635',
# '202249967',
# '803781764',
# '1723547798',
# '1726014416',
# '930686969',
# '1720080488',
# '927382275',
# '706047297',
# '302574678',
# '941646325',
# '1722232574',
# '930739685',
# '1205999038',
# '1754818019',
# '705917201',
# '705227171',
# '955367511',
# '2400009631',
# '921639290',
# '928930171',
# '926611120',
# '1205414277',
# '1207484609',
# '1315889764',
# '706668639',
# '1314511187',
# '604361139',
# '1105815672',
# '1105117327',
# '1313608430',
# '1205525346',
# '1726315060',
# '921847984',
# '1105060774',
# '250020864',
# '940286453',
# '1207027762',
# '2400090698',
# '944268754',
# '953694171',
# '301882569',
# '1207995299',
# '1724665235',
# '1206620047',
# '1723031611',
# '1719297150',
# '503444887',
# '1804467775',
# '401797865',
# '1206142851',
# '804098507',
# '604222117',
# '954699088',
# '2350706483',
# '914582523',
# '928985811',
# '705272078',
# '953483740',
# '2300717317',
# '1500794001',
# '1804546032',
# '1311969362',
# '1205322991',
# '803579325',
# '950472662',
# '105434070',
# '941646325',
# '1722232574',
# '104611348',
# '706724176',
# '1804625091',
# '803758747',
# '913506861',
# '1724462302',
# '929766541',
# '850161589',
# '2450018615',
# '1313115196',
# '930170535',
# '605114107',
# '1724133275',
# '1105079931',
# '703466326',
# '927707869',
# '1722191952',
# '922997663',
# '955342886',
# '1724832231',
# '921539219',
# '927398990',
# '706419447',
# '931074785',
# '1313650036',
# '107119562',
# '928882356',
# '1205521345',
# '1600600686',
# '1250148606',
# '202398053',
# '1314623883',
# '940337025',
# '1250222112',
# '941458911',
# '926333956',
# '1206651356',
# '1314075324',
# '1205311309',
# '929661106',
# '1105598252',
# '1719031575',
# '703466326',
# '801317686',
# '1722401591',
# '1314828631',
# '1105431215',
# '929894806',
# '706697869',
# '705327617',
# '1150243853',
# '502893415',
# '929187896',
# '1104938657',
# '930620588',
# '940301880',
# '1720042645',
# '926220930',
# '927501593',
# '926984709',
# '705418481',
# '1721929147',
# '931074785',
# '1313650036',
# '107119562',
# '1105535866',
# '104216924',
# '953415759',
# '803245786',
# '1719636969',
# '704468446',
# '927306456',
# '1104748221',
# '1726730722',
# '1723370019',
# '931369482',
# '953603941',
# '926333956',
# '1206651356',
# '929766541',
# '926776428',
# '921639290',
# '1105723579',
# '940752363',
# '2000093464',
# '202104584',
# '1500833312',
# '941150492',
# '1723971873',
# '1104596646',
# '604868406',
# '503385825',
# '919655274',
# '504121088',
# '1722401591',
# '1314828631',
# '1105431215',
# '1004282800',
# '930109707',
# '931999692',
# '2450282989',
# '928899103',
# '705836534',
# '927437202',
# '1207219492',
# '950392084',
# '930620588',
# '940301880',
# '1718873944',
# '927994442',
# '1314781574',
# '1003224951',
# '104779228',
# '1150166237',
# '707014049',
# '1104834922',
# '604761866',
# '1208128957',
# '1103847511',
# '605762459',
# '2100730346',
# '1600573552',
# '1805322532',
# '1311557464',
# '927993485',
# '706018488',
# '1204986986',
# '106838162',
# '951678077',
# '1207100452',
# '1207845536',
# '923468839',
# '604100610',
# '2100595277',
# '705187029',
# '503516163',
# '1207450253',
# '1205168592',
# '931640254',
# '1724996234',
# '929144681',
# '1205136714',
# '605555614',
# '1600750713',
# '1004224760',
# '1600643900',
# '2300554785',
# '302316203',
# '1500780737',
# '1205136714',
# '924216880',
# '750461576',
# '1207512623',
# '916470644',
# '1721981346',
# '802648915',
# '950561183',
# '302290994',
# '1722038203',
# '705172294',
# '1717531964',
# '1207922897',
# '958950933',
# '1312571639',
# '1724908536',
# '1205517509',
# '1721325528',
# '705717627',
# '503461170',
# '1207450253',
# '941197139',
# '105745038',
# '1722242144',
# '1106021635',
# '1315708378',
# '923742662',
# '1104968993',
# '953425972',
# '705860948',
# '503989055',
# '954571030',
# '705811651',
# '1206847293',
# '1205426206',
# '928421627',
# '951625441',
# '1723808844',
# '950476440',
# '750315806',
# '921337879',
# '805287745',
# '1207199447',
# '706694726',
# '804053346',
# '1723399166',
# '954236964',
# '922272638',
# '1750075002',
# '2350686966',
# '953221033',
# '958705469',
# '1313610303',
# '705398568',
# '1600813784',
# '1004282800',
# '930109707',
# '1723539787',
# '931509582',
# '927598409',
# '961117363',
# '1206712646',
# '706738515',
# '950414516',
# '950418418',
# '803070010',
# '1311167884',
# '1401145576',
# '929703700',
# '603864422',
# '1718436312',
# '1720669462',
# '1718873944',
# '927994442',
# '1314781574',
# '1310724040',
# '930432125',
# '1724135122',
# '706347499',
# '802648915',
# '1205136714',
# '605555614',
# '1724398530',
# '919336149',
# '1719345926',
# '929428951',
# '750278871',
# '1313441923',
# '925266066',
# '1311289621',
# '951678077',
# '1105077158',
# '1724928526',
# '1105122590',
# '1205136714',
# '924216880',
# '940226277',
# '926589607',
# '925703944',
# '957819378',
# '1207384569',
# '850401803',
# '1310782022',
# '953921988',
# '929300549',
# '803948660',
# '953425972',
# '1206696047',
# '604860296',
# '929182319',
# '705326981',
# '1725660656',
# '1205136714',
# '850401803',
# '202249967',
# '1104937741',
# '803874023',
# '302279138',
# '1723488654',
# '1718356429',
# '1004162960',
# '1722030226',
# '1724398530',
# '926215179',
# '926858739',
# '2300097710',
# '105370407',
# '705544625',
# '750278871',
# '803119940',
# '1313895631',
# '1205347667',
# '930885611',
# '955240098',
# '1207226299',
# '1500974520',
# '1313264283',
# '957051154',
# '940226277',
# '926589607',
# '925703944',
# '925232720',
# '929737930',
# '707341731',
# '1726725052',
# '952030971',
# '705363836',
# '1206696047',
# '850401803',
# '750466658',
# '401710876',
# '1207199447',
# '1723524680',
# '603038969',
# '1105571093',
# '956814768',
# '1206251546',
# '953689346',
# '1723613210',
# '930468905',
# '1726026121',
# '1205070566',
# '951887645',
# '1720423506',
# '1805525027',
# '1206041293',
# '706635711',
# '1104178452',
# '2300108368',
# '1206481333',
# '503741076',
# '1207455385',
# '930255153',
# '803119940',
# '706275229',
# '950288241',
# '1313335893',
# '1105473654',
# '955053863',
# '914802020',
# '706708856',
# '950452748',
# '925232720',
# '1718501883',
# '922801576',
# '1207211705',
# '925818601',
# '1726648957',
# '942272717',
# '953425972',
# '955851662',
# '930152756',
# '950645564',
# '1205475880',
# '917224982',
# '104485305',
# '1310438278',
# '1723022842',
# '1750206748',
# '1207995216',
# '1723990261',
# '929821411',
# '803119940',
# '929300549',
# '1104833965',
# '604156802',
# '2300456171',
# '1721185146',
# '928436302',
# '926444050',
# '1205929969',
# '704418300',
# '1724123243',
# '1719918243',
# '202041091',
# '956512073',
# '705726206',
# '202140828',
# '1725849143',
# '803119940',
# '917224982',
# '1313582304',
# '706413366',
# '940617558',
# '1206283879',
# '950763821',
# '1726648957',
# '925033870',
# '503461170',
# '1718486291',
# '503220139',
# '1718059304',
# '603437633',
# '1206029363',
# '930472386',
# '705254654',
# '923969489',
# '941420705',
# '1313582304',
# '504239989',
# '923768410',
# '917224982',
# '1121708687',
# '1105668527',
# '1206745844',
# '1312623307',
# '803151356',
# '940166465',
# '401361555',
# '950763821',
# '504348988',
# '1315445955',
# '927990051',
# '926205907',
# '1724021538',
# '1206448068',
# '106208226',
# '504092354',
# '1207242239',
# '1207470244',
# '1317134318',
# '941935470',
# '929738680',
# '1312139270',
# '927566588',
# '1205833765',
# '926920018',
# '706582152',
# '1721768958',
# '930170998',
# '927132472',
# '925693152',
# '1718695594',
# '928807494',
# '1150041893',
# '929916070',
# '930693072',
# '1723055784',
# '951887645',
# '1720423506',
# '1805525027',
# '1314845536',
# '921657409',
# '930268164',
# '941021826',
# '1313169136',
# '928876663',
# '851366500',
# '927566588',
# '1726749359',
# '957187685',
# '1501161846',
# '706627692',
# '1313772533',
# '604350603',
# '803463538',
# '1720570686',
# '922997663',
# '706927779',
# '931593982',
# '706275229',
# '950288241',
# '603931544',
# '927420190',
# '',
# '1720674686',
# '1206536458',
# '2100528211',
# '302151030',
# '1204924698',
# '104659834',
# '201948361',
# '1720444320',
# '1724412885',
# '923047807',
# '919374892',
# '105965453',
# '201601788',
# '604793166',
# '925538878',
# '502532948',
# '1500907447',
# '1717714750',
# '926003054',
# '920176773',
# '603805748',
# '929321784',
# '923977367',
# '922649033',
# '705602316',
# '1600488736',
# '1312140765',
# '1002588430',
# '922512397',
# '911359149',
# '704728401',
# '1717485153',
# '1104114846',
# '704524529',
# '920596335',
# '604665281',
# '503451262',
# '1802670990',
# '704512920',
# '604100610',
# '1204924698',
# '704709005',
# '1721654083',
# '1712257144',
# '105166185',
# '302151030',
# '2100279849',
# '502614860',
# '2300191299',
# '1206273524',
# '802519702',
# '1309934600',
# '1804095089',
# '703758698',
# '704236090',
# '1003757968',
# '105370407',
# '1205898099',
# '1204920589',
# '2100138375',
# '1004028377',
# '705165850',
# '1723054639',
# '802143800',
# '705034452',
# '1715711501',
# '1713944229',
# '201532850',
# '1310587736',
# '803258078',
# '1313210302',
# '1719257931',
# '921656336',
# '803053966',
# '927314518',
# '802927368',
# '1400452791',
# '803565761',
# '1206072280',
# '1103733794',
# '1312229923',
# '920159662',
# '1204578411',
# '917297707',
# '921557617',
# '1312062340',
# '1104431398',
# '803117688',
# '202284501',
# '802846840',
# '1312139270',
# '201956224',
# '705049013',
# '1204368342',
# '1311421653',
# '1723689863',
# '302098264',
# '922168737',
# '918023391',
# '927723346',
# '928920271',
# '923748123',
# '921540126',
# '1105063356',
# '603154857',
# '1713491874',
# '1724743073',
# '603381401',
# '401323258',
# '1723133995',
# '703753137',
# '929561033',
# '1720766409',
# '1003497714',
# '926165374',
# '1600546400',
# '704142355',
# '919433441',
# '1311650756',
# '931130454',
# '802864934',
# '1721258687',
# '2100138375',
# '917161630',
# '803602747',
# '920980596',
# '1105665614',
# '928989508',
# '1721388898',
# '920800471',
# '921971966',
# '401591110',
# '803258078',
# '1313210302',
# '1719257931',
# '921656336',
# '803053966',
# '927314518',
# '802927368',
# '1710315894',
# '201841178',
# '1715075949',
# '605063502',
# '604465732',
# '1308346046',
# '1804111688',
# '1400452791',
# '1103615264',
# '802215095',
# '925009607',
# '2200017917',
# '919268003',
# '704308428',
# '1311063331',
# '921287389',
# '1719966515',
# '927308155',
# '1719005579',
# '1600858292',
# '1900553213',
# '1900416965',
# '2400007494',
# '2100893920',
# '919337907',
# '1104598386',
# '705267235',
# '1206344358',
# '1103735062',
# '1718356429',
# '1720416179',
# '1717053621',
# '1400769855',
# '703861658',
# '604103259',
# '201876299',
# '803510213',
# '925955478',
# '1716550494',
# '602936650',
# '102493525',
# '604326041',
# '802295634',
# '502461965',
# '1713873675',
# '201552866',
# '503497489',
# '1715373948',
# '1716611098',
# '503640187',
# '502595523',
# '1308909264',
# '704303817',
# '103859211',
# '603377011',
# '1003524228',
# '604022285',
# '1718305004',
# '1720344686',
# '1720968187',
# '401539192',
# '914327663',
# '1311033698',
# '919321885',
# '301527966',
# '1714333984',
# '1711433993',
# '502461965',
# '1717148181',
# '1308909264',
# '928285212',
# '1803276524',
# '401484233',
# '704596733',
# '802961557',
# '1727180331',
# '201491644',
# '604103259',
# '605640267',
# '1500646102',
# '1723804165',
# '926886029',
# '923391494',
# '1721002044',
# '922665625',
# '1003705744',
# '1722692991',
# '1312938051',
# '1204771032',
# '1307418408',
# '1716310154',
# '919676858',
# '603281833',
# '924909930',
# '1400657720',
# '926068131',
# '1205919200',
# '1312326927',
# '502615883',
# '940352966',
# '1002754008',
# '1400452791',
# '1500891161',
# '1720416179',
# '1309057949',
# '917969909',
# '924183759',
# '917027344',
# '705262848',
# '927728519',
# '928362128',
# '1725220279',
# '704142355',
# '919433441',
# '1311650756',
# '931130454',
# '802864934',
# '1312679457',
# '706824711',
# '1715574487',
# '918010141',
# '927162172',
# '2300052764',
# '925752271',
# '804019412',
# '1723554208',
# '1205983099',
# '1720788817',
# '1721258687',
# '2100138375',
# '502999055',
# '1308940228',
# '503258907',
# '1600803330',
# '301931572',
# '401492327',
# '1716858392',
# '802748285',
# '924461981',
# '918566712',
# '201779915',
# '503620163',
# '1205640525',
# '502804420',
# '603161043',
# '1722401229',
# '1205480542',
# '924561160',
# '927010959',
# '926537762',
# '1103339832',
# '918737792',
# '929220879',
# '918185240',
# '923853634',
# '930839543',
# '930631437',
# '1804161535',
# '703772723',
# '1104930506',
# '1104177041',
# '603109323',
# '1711411635',
# '1310152457',
# '924909930',
# '801325788',
# '1311713828',
# '502417421',
# '1721688594',
# '1103764088',
# '2300272271',
# '931062103',
# '704890805',
# '924795271',
# '1204099707',
# '1718495748',
# '1204989311',
# '503049173',
# '104054754',
# '1727180331',
# '605352418',
# '803728211',
# '1720788817',
# '1310827397',
# '1717625394',
# '1206276469',
# '924110497',
# '503031742',
# '604936575',
# '604293050',
# '704165356',
# '1104557119',
# '1104165913',
# '1600542094',
# '1712424611',
# '201977345',
# '1002174637',
# '918118654',
# '106021371',
# '301572442',
# '919321885',
# '802520072',
# '925788614',
# '704912724',
# '911359149',
# '301699831',
# '1719250613',
# '104810700',
# '803243310',
# '1717583213',
# '917969909',
# '1206144832',
# '602866949',
# '929700078',
# '925622938',
# '924141104',
# '921904215',
# '1804234290',
# '1900579325',
# '1500912256',
# '104054754',
# '1716038680',
# '1722776943',
# '1310356603',
# '803084607',
# '926254830',
# '1804013470',
# '704003789',
# '1307418408',
# '201569357',
# '924306004',
# '401842828',
# '503010175',
# '1206203836',
# '929821411',
# '104054754',
# '1727180331',
# '605352418',
# '922579636',
# '924110497',
# '1803710522',
# '1803193042',
# '302059621',
# '1803235074',
# '1723055958',
# '103258521',
# '1103941066',
# '704403559',
# '921030698',
# '803139088',
# '1204217218',
# '704825256',
# '924186661',
# '1205089251',
# '1207464551',
# '104863584',
# '1104335409',
# '922060272',
# '923838148',
# '925737231',
# '926807553',
# '915871073',
# '930621735',
# '704732940',
# '919023382',
# '920999786',
# '1715203327',
# '1002578837',
# '803168491',
# '919524983',
# '1206093054',
# '919652966',
# '940394182',
# '704633536',
# '1715340087',
# '1104027832',
# '919922161',
# '603883182',
# '1103878458',
# '1719177378',
# '503859837',
# '1719298489',
# '1206180091',
# '602923450',
# '604353201',
# '704647734',
# '1308519600',
# '1103535645',
# '1104666597',
# '1103698252',
# '1803710522',
# '1803193042',
# '704344688',
# '925557712',
# '1309178133',
# '1722392873',
# '926414962',
# '926938259',
# '105326532',
# '1712290228',
# '1206434084',
# '201956224',
# '1003082474',
# '802925610',
# '1713057618',
# '1310587736',
# '917297707',
# '930706528',
# '1205848367',
# '1718356429',
# '1205590423',
# '705398428',
# '1104560782',
# '603981622',
# '1204208506',
# '1725106064',
# '1350696124',
# '302253935',
# '1719728634',
# '1724432271',
# '1500444086',
# '926384371',
# '1400548507',
# '1204629198',
# '1205542879',
# '704647734',
# '1003594908',
# '1722526082',
# '920203098',
# '1803783545',
# '919896720',
# '2200084503',
# '1751537885',
# '925052698',
# '919465062',
# '802958272',
# '920961968',
# '1208629764',
# '1720157310',
# '1501005266',
# '925503369',
# '1722393236',
# '704712140',
# '604607416',
# '1003072913',
# '918928706',
# '2100138375',
# '1715538813',
# '201601010',
# '922537980',
# '706029469',
# '1500636889',
# '1205072067',
# '1802569119',
# '1104554827',
# '1720325487',
# '1206434084',
# '201956224',
# '1003082474',
# '202027140',
# '604393033',
# '927824771',
# '1721807541',
# '925547978',
# '1723055958',
# '929300580',
# '1206343483',
# '923600613',
# '201801313',
# '921390928',
# '918181546',
# '1717665721',
# '704888460',
# '921038469',
# '1600493611',
# '802956185',
# '1400663025',
# '1309592671',
# '2200084503',
# '921337754',
# '1712053840',
# '916804644',
# '920660123',
# '704272954',
# '919254144',
# '2200211957',
# '603347519',
# '1204207037',
# '918593849',
# '1724131196',
# '1727329714',
# '703549139',
# '802958272',
# '1311045064',
# '927734632',
# '921416475',
# '202324190',
# '1313757625',
# '2100371653',
# '1206013342',
# '1104212517',
# '103618674',
# '916489446',
# '1722281217',
# '1206255752',
# '502221807',
# '602992422',
# '803060128',
# '603977513',
# '1714829163',
# '704807593',
# '1724391055',
# '1713754289',
# '705095933',
# '603376310',
# '1206136663',
# '1714640982',
# '604253302',
# '1309588208',
# '504308933',
# '1204680670',
# '1314748441',
# '918181546',
# '928508183',
# '1206510784',
# '705348290',
# '1207096296',
# '951431758',
# '1204949687',
# '1003003033',
# '1206045021',
# '1900396241',
# '502801749',
# '922330345',
# '705326841',
# '1718791856',
# '925721623',
# '1103651855',
# '1718570789',
# '1723523138',
# '1721707212',
# '920543360',
# '921562088',
# '927429050',
# '1804013470',
# '1721383204',
# '1724095656',
# '1900494756',
# '1717760704',
# '919724526',
# '929705887',
# '922748629',
# '1719353110',
# '931191415',
# '1900441971',
# '919591388',
# '929370690',
# '1500978026',
# '921022075',
# '803791821',
# '929390458',
# '1900494756',
# '925752909',
# '1103748909',
# '923748495',
# '2100989587',
# '924239478',
# '923744163',
# '924183551',
# '1804580528',
# '301526133',
# '1600357303',
# '705811164',
# '802339531',
# '1720055167',
# '917830259',
# '2300146491',
# '925830911',
# '1720136637',
# '916700917',
# '920498938',
# '1725189086',
# '1204577702',
# '802749580',
# '930131115',
# '928645340',
# '802795450',
# '920660123',
# '704481373',
# '1205240664',
# '202324190',
# '706396132',
# '918864216',
# '919559757',
# '604495416',
# '941506925',
# '1715537989',
# '1803080256',
# '703165118',
# '1717324238',
# '926308164',
# '919254144',
# '1205089228',
# '1202975965',
# '2300549017',
# '2100233796',
# '1725441040',
# '1720638269',
# '705079010',
# '1104704414',
# '201832565',
# '1900517465',
# '921257804',
# '1105433567',
# '1002924668',
# '201700143',
# '1804220240',
# '201464401',
# '705095933',
# '919559757',
# '1718939133',
# '1202975965',
# '1716006638',
# '502473309',
# '914327663',
# '1203905425',
# '1311420739',
# '1310226780',
# '1725488397',
# '921337754',
# '604142760',
# '1718251083',
# '1206295188',
# '104198452',
# '1717281719',
# '917350613',
# '929456044',
# '1715144927',
# '603550740',
# '1205465923',
# '202324190',
# '802749580',
# '604406298',
# '920373560',
# '1725029381',
# '1717241390',
# '802003780',
# '2300549017',
# '2100233796',
# '1725441040',
# '1720638269',
# '1713297586',
# '925250276',
# '1725849143',
# '1002924668',
# '201700143',
# '1718251083',
# '2100221312',
# '1202975965',
# '502473309',
# '914327663',
# '1724111149',
# '1311420739',
# '1310226780',
# '1725488397',
# '929182319',
# '503469777',
# '1500770498',
# '1724044670',
# '924264005',
# '801920505',
# '908229594',
# '705699163',
# '2100233796',
# '603297094',
# '918772526',
# '1205956863',
# '925693152',
# '1716325657',
# '503054124',
# '919455980',
# '1207461573',
# '1716321987',
# '603454703',
# '1104426711',
# '1312024076',
# '604831214',
# '1203905425',
# '201644994',
# '1309163150',
# '603159815',
# '2100339924',
# '504089640',
# '704482066',
# '802331306',
# '1205789090',
# '1500958804',
# '915898928',
# '803492164',
# '923339881',
# '920792975',
# '927417113',
# '703916247',
# '1400718993',
# '801920505',
# '1205964826',
# '604882332',
# '1900441971',
# '703865212',
# '1205223686',
# '704464114',
# '930535695',
# '919521260',
# '604150490',
# '703078386',
# '1203793557',
# '1105621922',
# '1104525645',
# '104668173',
# '925044430',
# '1205730953',
# '1717310427',
# '1206723734',
# '925693152',
# '503054124',
# '2100279849',
# '915855266',
# '917020539',
# '1724904592',
# '1714709779',
# '1203905425',
# '201644994',
# '1312215393',
# '704728401',
# '104702857',
# '1712780491',
# '704723717',
# '1722730494',
# '103428249',
# '1003568837',
# '2200277818',
# '705957363',
# '1105077687',
# '1105621922',
# '926603598',
# '1309817169',
# '1716274731',
# '2100297106',
# '1003164348',
# '106513260',
# '704518349',
# '1720863677',
# '1721802716',
# '1723311534',
# '1717754269',
# '803903046',
# '927776708',
# '1104063845',
# '704482066',
# '917453367',
# '1206713800',
# '918532771',
# '923253587',
# '1722273644',
# '920060712',
# '1718394941',
# '201493855',
# '1724091671',
# '1723327415',
# '1717930802',
# '1207496850',
# '1600568198',
# '1720666906',
# '1003594908',
# '1205138066',
# '1718532433',
# '1309085833',
# '604151894',
# '502940349',
# '1500900533',
# '502287675',
# '920769973',
# '924306004',
# '703916247',
# '1717965568',
# '1309817169',
# '1312324997',
# '705048056',
# '801828732',
# '1205964826',
# '1104538432',
# '604882332',
# '503772055',
# '1103395834',
# '401192216',
# '1312215393',
# '803566462',
# '105188379',
# '925050494',
# '940125602',
# '921739314',
# '1204572604',
# '941922163',
# '914695333',
# '202100970',
# '302556568',
# '1150267639',
# '704733492',
# '1312346461',
# '1207496850',
# '1309085833',
# '940187313',
# '1717965568',
# '1309817169',
# '502939515',
# '603219452',
# '704790997',
# '704536937',
# '1104538432',
# '106021371',
# '2100493028',
# '940901721',
# '704851849',
# '940617558',
# '931590038',
# '1724716715',
# '940125602',
# '930497417',
# '1720666906',
# '603945841',
# '928002534',
# '302556568',
# '917982449',
# '925355943',
# '1721313631',
# '703412197',
# '301756151',
# '1003903349',
# '922935770',
# '1715572267',
# '106021371',
# '602991614',
# '924620768',
# '603848946',
# '1717282287',
# '922618145',
# '1716465081',
# '1724348782',
# '605185271',
# '202081196',
# '925326076',
# '1104178601',
# '920630563',
# '918825621',
# '1500601594',
# '1721313631',
# '1714640982',
# '1716476302',
# '922419965',
# '931508808',
# '917557498',
# '302344163',
# '503979023',
# '927207985',
# '1717282287',
# '922618145',
# '926594391',
# '804234540',
# '1714308549',
# '924620768',
# '1726207549',
# '921864674',
# '917457137',
# '503514499',
# '604615278',
# '1311527046',
# '923227201',
# '2300889975',
# '2100989587',
# '921347688',
# '1204395584',
# '605160118',
# '104659834',
# '1720712502',
# '1715136147',
# '603383605',
# '401466099',
# '1500552110',
# '1708013444',
# '924609654',
# '201869476',
# '801559469',
# '1002174637',
# '603383605',
# '1714640982',
# '502801749',
# '929775252',
# '603390105',
# '1309508388',
# '1206369207',
# '1721323382',
# '927223529',
# '931665186',
# '929258036',
# '1804004644',
# '1205871559',
# '924536212',
# '1204578411',
# '1104926264',
# '1717330953',
# '1205796673',
# '919537514',
# '1722341946',
# '1600585135',
# '917557498',
# '925083628',
# '1600468720',
# '802806836',
# '503675407',
# '926594391',
# '908652324',
# '103258521',
# '604022285',
# '1720836467',
# '1104010184',
# '504308933',
# '940493588',
# '1204176083',
# '1206482307',
# '1103533418',
# '1804260873',
# '923702641',
# '105188379',
# '1104380660',
# '1714158886',
# '301584579',
# '201470960',
# '1725849143',
# '923048995',
# '1804250031',
# '502650468',
# '503159535',
# '927571331',
# '924421241',
# '104659834',
# '502650468',
# '914412945',
# '103667309',
# '1202975965',
# '1803467370',
# '921303897',
# '2100428230',
# '1724907520',
# '1717928806',
# '401745153',
# '704560580',
# '201647781',
# '706126760',
# '922675236',
# '1721267571',
# '704541770',
# '1717952301',
# '913674842',
# '1312863093',
# '502801749',
# '917961484',
# '1725107641',
# '930123062',
# '923594964',
# '1206097618',
# '921406666',
# '926091570',
# '1716273725',
# '926335811',
# '919945238',
# '924366396',
# '1715164362',
# '2300189749',
# '1207619584',
# '1900669266',
# '940856305',
# '1312887076',
# '920854429',
# '924290315',
# '925083628',
# '1714003223',
# '603597238',
# '920336773',
# '1721049326',
# '1205639972',
# '705049435',
# '1207807320',
# '1803780061',
# '603291055',
# '1207488261',
# '916296528',
# '923827653',
# '930512538',
# '926335811',
# '1900760719',
# '1804213120',
# '1804019618',
# '1722567839',
# '401344536',
# '1103985469',
# '1757809759',
# '1711105971',
# '1251221659',
# '302098264',
# '917380305',
# '1104380660',
# '1714158886',
# '301584579',
# '201470960',
# '1003696117',
# '920057924',
# '1105023905',
# '105229785',
# '1204576126',
# '1104290349',
# '1104293178',
# '1309817169',
# '1719236505',
# '802224543',
# '1003696117',
# '920057924',
# '1105023905',
# '105229785',
# '1204576126',
# '1104290349',
# '1104293178',
# '1309817169',
# '1719236505',
# '802224543',
# '1803710522',
# '1802155620',
# '604196584',
# '202149852',
# '703886721',
# '1714362785',
# '924329600',
# '914412945',
# '503318826',
# '105378715',
# '1003249198',
# '1103753552',
# '706126760',
# '922675236',
# '503765059',
# '1714590484',
# '302433594',
# '1316676913',
# '703559435',
# '1900617455',
# '1104974678',
# '1804744983',
# '703339572',
# '302211966',
# '1724432651',
# '704102482',
# '924527690',
# '704886340',
# '1314011634',
# '401253323',
# '919945238',
# '1722564638',
# '604077313',
# '1204647844',
# '1205160425',
# '1717275562',
# '1207619584',
# '1719749069',
# '953839578',
# '801851825',
# '920908233',
# '926872466',
# '925349359',
# '1757809759',
# '2100274857',
# '1714003223',
# '603597238',
# '920336773',
# '1721049326',
# '1103615264',
# '1201693742',
# '1803780061',
# '603291055',
# '1207488261',
# '916296528',
# '302433594',
# '923368534',
# '1716330343',
# '301815338',
# '1720690187',
# '1104950066',
# '503530982',
# '705386514',
# '2400007098',
# '920084126',
# '1003696117',
# '920057924',
# '1105023905',
# '1724124043',
# '919095182',
# '1310221765',
# '301844700',
# '1002851192',
# '923639363',
# '1900760719',
# '1104380660',
# '1803710522',
# '1206055483',
# '1715664072',
# '604095000',
# '926073834',
# '604196584',
# '1002696712',
# '917646200',
# '1720787918',
# '1314825496',
# '1722997341',
# '915076558',
# '1311460602',
# '1724962582',
# '1804087581',
# '503765059',
# '1714590484',
# '302433594',
# '503807802',
# '1204584187',
# '704509660',
# '922426184',
# '1105629123',
# '1309878138',
# '1204538910',
# '1722564638',
# '1720292372',
# '202027140',
# '926304569',
# '1003175450',
# '1720048527',
# '1207510320',
# '1205375049',
# '1713584678',
# '802124586',
# '802634782',
# '1721807541',
# '1206880617',
# '940764327',
# '1715192686',
# '1104003619',
# '301092714',
# '1720787918',
# '925788614',
# '802950238',
# '302433594',
# '102493525',
# '1717348294',
# '1206180091',
# '926177445',
# '1104676257',
# '302025622',
# '1104527021',
# '1715942346',
# '1003194121',
# '1308804309',
# '201289113',
# '201918679',
# '1309850509',
# '924230949',
# '1719572669',
# '503461279',
# '202098729',
# '704729755',
# '1724365919',
# '917322018',
# '704549658',
# '918759283',
# '1313243493',
# '1105629123',
# '1206485474',
# '918855651',
# '1722356753',
# '930146147',
# '1721474003',
# '1105242075',
# '2350658445',
# '929751170',
# '918619214',
# '1713801767',
# '940968639',
# '926352451',
# '923963698',
# '919101816',
# '604831214',
# '1715192686',
# '201745171',
# '603396839',
# '1206723411',
# '302100565',
# '1718230822',
# '929371581',
# '921660973',
# '1719297150',
# '201879061',
# '1205461773',
# '1500963739',
# '1722209002',
# '925977084',
# '924916067',
# '803478395',
# '1103925598',
# '2100982988',
# '1206413419',
# '924265036',
# '1204681264',
# '923963698',
# '914797261',
# '603879982',
# '1712441029',
# '104655683',
# '1309809190',
# '921653234',
# '1205100512',
# '951799899',
# '1720325487',
# '201815156',
# '916489446',
# '103386777',
# '704541481',
# '1103738702',
# '1205824079',
# '1716188345',
# '102493525',
# '1717348294',
# '1714208848',
# '1714074687',
# '925083016',
# '603358854',
# '104655683',
# '201564044',
# '603091414',
# '1803185931',
# '1724134489',
# '603399528',
# '502226673',
# '920687472',
# '921937520',
# '503077075',
# '1715685952',
# '1718207069',
# '1311626616',
# '1718495300',
# '502678170',
# '201664299',
# '916762859',
# '925497521',
# '603981895',
# '1712441029',
# '923848493',
# '1003020003',
# '1720147154',
# '201289113',
# '1003242136',
# '502226673',
# '919699421',
# '1720866639',
# '1310957129',
# '706273646',
# '1802930758',
# '921482865',
# '1104198211',
# '107083370',
# '930037460',
# '1709971087',
# '503858748',
# '1310303209',
# '804135218',
# '923392625',
# '920866191',
# '1206659581',
# '921495727',
# '923601934',
# '923969869',
# '1311021511',
# '924091804',
# '1722269105',
# '930877931',
# '2100494653',
# '930511076',
# '1205375049',
# '1308645504',
# '940800527',
# '802705558',
# '703736330',
# '1204152373',
# '1723426662',
# '914581822',
# '940702418',
# '1311538217',
# '1500751688',
# '1719572669',
# '922179783',
# '802514927',
# '915740732',
# '1205704883',
# '919310607',
# '918716127',
# '918675133',
# '703916247',
# '922016605',
# '922804703',
# '1723636690',
# '803034446',
# '702974320',
# '941225534',
# '922752183',
# '504072448',
# '803500891',
# '921592242',
# '1314067396',
# '941068488',
# '921130712',
# '921500351',
# '1717174450',
# '803503242',
# '604125351',
# '602869810',
# '1717189193',
# '920529922',
# '929562395',
# '1400630214',
# '918759283',
# '705147544',
# '105480453',
# '1723042436',
# '105485130',
# '919147876',
# '925622961',
# '930133236',
# '1206485474',
# '918855651',
# '1722356753',
# '704643865',
# '921347324',
# '920866191',
# '924091028',
# '1311139768',
# '925497521',
# '1204879553',
# '924535727',
# '604125351',
# '1309920278',
# '919972752',
# '924823529',
# '401585112',
# '1205177262',
# '704549658',
# '2350658445',
# '915937114',
# '923969836',
# '2200084503',
# '1717562654',
# '202007100',
# '1718799917',
# '951268432',
# '706031671',
# '922179783',
# '919101816',
# ])
#
#         filas_recorridas = 4
#         cont = 1
#
#         for postulante in postulantes:
#             if postulante.tiene_matricula_cohorte():
#                 if postulante.retirado_matricula():
#                     estado = 'Retirado'
#                 else:
#                     estado = 'Matriculado'
#             elif postulante.inscripcion:
#                 estado = 'Inscrito'
#             elif postulante.estado_aprobador == 1:
#                 estado = 'En proceso'
#             elif postulante.estado_aprobador == 2:
#                 estado = 'Admitido'
#             elif postulante.estado_aprobador == 3:
#                 estado = 'Rechazado'
#
#             ws.write('A%s' % filas_recorridas, str(cont), formatoceldaleft)
#             ws.write('B%s' % filas_recorridas, str(postulante.id), formatoceldaleft)
#             ws.write('C%s' % filas_recorridas, str(postulante.inscripcionaspirante.persona.identificacion()), formatoceldaleft)
#             ws.write('D%s' % filas_recorridas, str(postulante.inscripcionaspirante.persona.nombre_completo_inverso()), formatoceldaleft)
#             ws.write('E%s' % filas_recorridas, str(postulante.cohortes.maestriaadmision.descripcion), formatoceldaleft)
#             ws.write('F%s' % filas_recorridas, str(postulante.cohortes.descripcion), formatoceldaleft)
#             ws.write('G%s' % filas_recorridas, str(postulante.fecha_creacion.date()), formatoceldaleft)
#             ws.write('H%s' % filas_recorridas, str(estado), formatoceldaleft)
#             ws.write('I%s' % filas_recorridas, str('Si' if postulante.status else 'No'), formatoceldaleft)
#
#             filas_recorridas += 1
#             cont += 1
#
#         ws = workbook.add_worksheet('QUITO')
#         ws.set_column(0, 0, 10)
#         ws.set_column(1, 1, 15)
#         ws.set_column(2, 2, 15)
#         ws.set_column(3, 3, 30)
#         ws.set_column(4, 4, 30)
#         ws.set_column(5, 5, 40)
#         ws.set_column(6, 6, 30)
#         ws.set_column(7, 7, 30)
#
#         ws.write(2, 0, 'N°', formatoceldacab)
#         ws.write(2, 1, 'Id', formatoceldacab)
#         ws.write(2, 2, 'Cédula', formatoceldacab)
#         ws.write(2, 3, 'Postulante', formatoceldacab)
#         ws.write(2, 4, 'Maestría', formatoceldacab)
#         ws.write(2, 5, 'Cohorte', formatoceldacab)
#         ws.write(2, 6, 'Fecha de postulación', formatoceldacab)
#         ws.write(2, 7, 'Estado', formatoceldacab)
#         ws.write(2, 8, '¿Activo?', formatoceldacab)
#
#         postulantes = InscripcionCohorte.objects.filter(inscripcionaspirante__persona__cedula__in=['0503354789',
# '0604568162',
# '2450190877',
# '1711006872',
# '1716821853',
# '1719322545',
# '0914783832',
# '0401484233',
# '0502940174',
# '0703078386',
# '0913338562',
# '0703078386',
# '1711336766',
# '1716655632',
# '0201813813',
# '0912297983',
# '1720416179',
# '0907877476',
# '0912297983',
# '1707242853',
# '0805009875',
# '1709697807',
# '0502428220',
# '1713567327',
# '1900257211',
# '0924461981',
# '0918566712',
# '0201779915',
# '0503620163',
# '1205640525',
# ])
#
#         filas_recorridas = 4
#         cont = 1
#
#         for postulante in postulantes:
#             if postulante.tiene_matricula_cohorte():
#                 if postulante.retirado_matricula():
#                     estado = 'Retirado'
#                 else:
#                     estado = 'Matriculado'
#             elif postulante.inscripcion:
#                 estado = 'Inscrito'
#             elif postulante.estado_aprobador == 1:
#                 estado = 'En proceso'
#             elif postulante.estado_aprobador == 2:
#                 estado = 'Admitido'
#             elif postulante.estado_aprobador == 3:
#                 estado = 'Rechazado'
#
#             ws.write('A%s' % filas_recorridas, str(cont), formatoceldaleft)
#             ws.write('B%s' % filas_recorridas, str(postulante.id), formatoceldaleft)
#             ws.write('C%s' % filas_recorridas, str(postulante.inscripcionaspirante.persona.identificacion()), formatoceldaleft)
#             ws.write('D%s' % filas_recorridas, str(postulante.inscripcionaspirante.persona.nombre_completo_inverso()), formatoceldaleft)
#             ws.write('E%s' % filas_recorridas, str(postulante.cohortes.maestriaadmision.descripcion), formatoceldaleft)
#             ws.write('F%s' % filas_recorridas, str(postulante.cohortes.descripcion), formatoceldaleft)
#             ws.write('G%s' % filas_recorridas, str(postulante.fecha_creacion.date()), formatoceldaleft)
#             ws.write('H%s' % filas_recorridas, str(estado), formatoceldaleft)
#             ws.write('I%s' % filas_recorridas, str('Si' if postulante.status else 'No'), formatoceldaleft)
#
#             filas_recorridas += 1
#             cont += 1
#
#         ws = workbook.add_worksheet('MANTA')
#         ws.set_column(0, 0, 10)
#         ws.set_column(1, 1, 15)
#         ws.set_column(2, 2, 15)
#         ws.set_column(3, 3, 30)
#         ws.set_column(4, 4, 30)
#         ws.set_column(5, 5, 40)
#         ws.set_column(6, 6, 30)
#         ws.set_column(7, 7, 30)
#
#         ws.write(2, 0, 'N°', formatoceldacab)
#         ws.write(2, 1, 'Id', formatoceldacab)
#         ws.write(2, 2, 'Cédula', formatoceldacab)
#         ws.write(2, 3, 'Postulante', formatoceldacab)
#         ws.write(2, 4, 'Maestría', formatoceldacab)
#         ws.write(2, 5, 'Cohorte', formatoceldacab)
#         ws.write(2, 6, 'Fecha de postulación', formatoceldacab)
#         ws.write(2, 7, 'Estado', formatoceldacab)
#         ws.write(2, 8, '¿Activo?', formatoceldacab)
#
#         postulantes = InscripcionCohorte.objects.filter(inscripcionaspirante__persona__cedula__in=['0929293504',
# '1205288317',
# '1207268937',
# '1204892754',
# '1312113291',
# '1311289621',
# '0930851969',
# '1207527944',
# '1206502963',
# '1724904592',
# '1719266379',
# '1715172050',
# '1351426869',
# '1208603595',
# '0804320158',
# '0929293504',
# '1205288317',
# '1207268937',
# '1204892754',
# '1050094810',
# '1722675772',
# '1722038203',
# '0925892226',
# '1310587736',
# '1721383204',
# '1205310517',
# '1050094810',
# '0803443969',
# '1718395286',
# '0803443969',
# '1351645872',
# '0803408871',
# '2300448095',
# '0942015413',
# ])
#
#         filas_recorridas = 4
#         cont = 1
#
#         for postulante in postulantes:
#             if postulante.tiene_matricula_cohorte():
#                 if postulante.retirado_matricula():
#                     estado = 'Retirado'
#                 else:
#                     estado = 'Matriculado'
#             elif postulante.inscripcion:
#                 estado = 'Inscrito'
#             elif postulante.estado_aprobador == 1:
#                 estado = 'En proceso'
#             elif postulante.estado_aprobador == 2:
#                 estado = 'Admitido'
#             elif postulante.estado_aprobador == 3:
#                 estado = 'Rechazado'
#
#             ws.write('A%s' % filas_recorridas, str(cont), formatoceldaleft)
#             ws.write('B%s' % filas_recorridas, str(postulante.id), formatoceldaleft)
#             ws.write('C%s' % filas_recorridas, str(postulante.inscripcionaspirante.persona.identificacion()), formatoceldaleft)
#             ws.write('D%s' % filas_recorridas, str(postulante.inscripcionaspirante.persona.nombre_completo_inverso()), formatoceldaleft)
#             ws.write('E%s' % filas_recorridas, str(postulante.cohortes.maestriaadmision.descripcion), formatoceldaleft)
#             ws.write('F%s' % filas_recorridas, str(postulante.cohortes.descripcion), formatoceldaleft)
#             ws.write('G%s' % filas_recorridas, str(postulante.fecha_creacion.date()), formatoceldaleft)
#             ws.write('H%s' % filas_recorridas, str(estado), formatoceldaleft)
#             ws.write('I%s' % filas_recorridas, str('Si' if postulante.status else 'No'), formatoceldaleft)
#
#             filas_recorridas += 1
#             cont += 1
#
#         workbook.close()
#         # output.seek(0)
#         # fecha_hora_actual = datetime.now().date()
#         # filename = 'Listado_ventas_' + str(fecha_hora_actual) + '.xlsx'
#         response = HttpResponse(directory, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#         response['Content-Disposition'] = 'attachment; filename=%s' % name_document
#         url_file = "{}reportes/gestion/{}".format(MEDIA_URL, nombre_archivo)
#         print(url_file)
#     except Exception as ex:
#         pass
#
# imprimirreporte()
    # print("Migracion de leads de asesores antiguos Merino-Cruz")
    # antiguo = AsesorComercial.objects.get(status=True, pk=35)
    # nuevo = AsesorComercial.objects.get(status=True, pk=38)
    # idventas = VentasProgramaMaestria.objects.filter(status=True, asesor=antiguo).values_list('inscripcioncohorte__id', flat=True)
    # leads = InscripcionCohorte.objects.filter(status=True, asesor=antiguo, vendido=False).exclude(id__in=idventas)
    # c=0
    # for lead in leads:
    #     lead.asesor = nuevo
    #     lead.tiporespuesta = None
    #     lead.save()
    #
    #     histoanti = HistorialAsesor.objects.get(inscripcion_id=lead.id, fecha_fin=None)
    #     histoanti.fecha_fin = lead.fecha_modificacion
    #     histoanti.save()
    #     histo = HistorialAsesor(inscripcion_id=lead.id, fecha_inicio=lead.fecha_modificacion,
    #                             fecha_fin=None, asesor=lead.asesor, observacion='MASIVO - LEAD DE ASESOR INACTIVO')
    #     histo.save()
    #     c += 1
    #     print(f'{c} - Lead: {lead} - asesor {lead.asesor}')
    #
    # print("Migracion de leads de asesores antiguos Morocho-Hurtado")
    # antiguo = AsesorComercial.objects.get(status=True, pk=36)
    # nuevo = AsesorComercial.objects.get(status=True, pk=1)
    # idventas = VentasProgramaMaestria.objects.filter(status=True, asesor=antiguo).values_list('inscripcioncohorte__id', flat=True)
    # leads = InscripcionCohorte.objects.filter(status=True, asesor=antiguo, vendido=False).exclude(id__in=idventas)
    # c=0
    # for lead in leads:
    #     lead.asesor = nuevo
    #     lead.tiporespuesta = None
    #     lead.save()
    #
    #     histoanti = HistorialAsesor.objects.get(inscripcion_id=lead.id, fecha_fin=None)
    #     histoanti.fecha_fin = lead.fecha_modificacion
    #     histoanti.save()
    #     histo = HistorialAsesor(inscripcion_id=lead.id, fecha_inicio=lead.fecha_modificacion,
    #                             fecha_fin=None, asesor=lead.asesor, observacion='MASIVO - LEAD DE ASESOR INACTIVO')
    #     histo.save()
    #     c += 1
    #     print(f'{c} - Lead: {lead} - asesor {lead.asesor}')
    #
#     print("Migracion de leads de contabilidad")
#
#     cohorte2023 = CohorteMaestria.objects.get(id=173, status=True)
#     cont = 0
#     matriculados = []
#     fvence = datetime.strptime('2023-08-20', '%Y-%m-%d').date()
#
#     postu = InscripcionCohorte.objects.filter(status=True, cohortes__id=173)
#     for pos in postu:
#         if pos.tiene_matricula_cohorte():
#             matriculados.append(pos.id)
#
#     postulantes = InscripcionCohorte.objects.filter(estado_aprobador__in=[1, 2], cohortes__id=173, id__in=[40927,
# 24432,
# 52095,
# 28902,
# 44155,
# 53998,
# 62225,
# 29845,
# 33277,
# 55254,
# 51213,
# 52569,
# 53130,
# 55278,
# 58843,
# 54975,
# 45196,
# 56783,
# 24831,
# 58856,
# 43843,
# 61242,
# 50707,
# 21696,
# 48453,
# 55023,
# 23475,
# 55297,
# 22098,
# 39161,
# 21843,
# 55374,
# 62271,
# 56323,
# 56322,
# 22043,
# 22069,
# 49510,
# 22097,
# 52407,
# 60995,
# 21874,
# 22260,
# 22656,
# 22935,
# 50565,
# 22732,
# 33604,
# 21971,
# 22463,
# 22975,
# 22473,
# 31704,
# 50571,
# 56367,
# 63819,
# 34027,
# 31839,
# 63822,
# 39237,
# 33536,
# 32495,
# 52182,
# 22918,
# 64368,
# 36840,
# 36325,
# 36337,
# 36700,
# 32694,
# 36788,
# 52426,
# 49113,
# 52189,
# 46922,
# 49624,
# 49630,
# 56976,
# 37404,
# 35767,
# 21915,
# 37235,
# 27694,
# 43571,
# 52445,
# 58839,
# 31186,
# 45391,
# 22902,
# 21925,
# 34522,
# 57336,
# 57317,
# 39428,
# 21934,
# 26408,
# 35302,
# 37364,
# 37681,
# 49519,
# 56366,
# 33409,
# 26772,
# 38272,
# 56287,
# 22208,
# 22766,
# 52273,
# 33513,
# 53683,
# 22756,
# 45797,
# 36062,
# 53739,
# 42727,
# 23373,
# 27733,
# 34302,
# 34347,
# 28246,
# 34455,
# 44652,
# 49643,
# 57179,
# 21942,
# 44226,
# 22247,
# 27352,
# 21784,
# 29078,
# 21919,
# 61918,
# 24208,
# 37198,
# 40664,
# 42420,
# 21717,
# 31182,
# 55750,
# 44331,
# 43175,
# 44888,
# 34657,
# 31835,
# 56237,
# 52715,
# 58680,
# 32476,
# 56303,
# 32522,
# 55936,
# 34943,
# 34963,
# 58737,
# 41219,
# 50232,
# 60960,
# 36558,
# 52097,
# 52135,
# 52155,
# 52303,
# 53085,
# 54211,
# 54227,
# 54748,
# 54756,
# 54886,
# 54926,
# 55009,
# 55010,
# 55027,
# 55143,
# 39845,
# 56263,
# 56500,
# 56544,
# 42123,
# 34795,
# 53983,
# 44894,
# 48387,
# 55506,
# 58938,
# 58914,
# 58960,
# 60840,
# 60849,
# 61410,
# 61778,
# 35733,
# 53981,
# 54717,
# 56167,
# 59881,
# 54950,
# 55144,
# 44859,
# 55564,
# 61680,
# 59349,
# 58736,
# 59973,
# 58263,
# 44508,
# 60004,
# 46950,
# 60713,
# 53289,
# 57987,
# 60576,
# 62924,
# 60973,
# 36535,
# 62968,
# 63414,
# 63029,
# 48878,
# 40012,
# 49201,
# 63091,
# 57145,
# 40218,
# 37678,
# 61542,
# 50143,
# 50275,
# 50298,
# 50308,
# 50452,
# 50689,
# 50842,
# 51237,
# 52479,
# 52894,
# 64948,
# 54313,
# 54653,
# 54738,
# 44301,
# 65071,
# 55081,
# 55087,
# 65335,
# 65596,
# 57140,
# 59511,
# 61080,
# 61545,
# 64832,
# 65836,
# 65856,
# 22145,
# 22156,
# 22196,
# 22968,
# 24269,
# 24459,
# 25996,
# 26065,
# 26729,
# 26883,
# 27410,
# 27416,
# 38066,
# 50436,
# 52616,
# 53014,
# 55864,
# 55866,
# 63617,
# 63355,
# 21650,
# 21687,
# 21688,
# 21710,
# 21728,
# 21792,
# 21812,
# 21821,
# 21831,
# 21832,
# 21859,
# 21867,
# 21869,
# 21873,
# 21883,
# 21900,
# 21902,
# 21912,
# 21913,
# 21918,
# 21929,
# 21931,
# 21939,
# 21947,
# 21964,
# 21969,
# 21981,
# 21985,
# 21987,
# 22013,
# 22036,
# 22122,
# 22126,
# 22238,
# 22269,
# 22283,
# 22356,
# 22413,
# 22607,
# 22710,
# 22712,
# 23021,
# 23025,
# 23029,
# 23250,
# 23379,
# 23452,
# 23478,
# 23580,
# 23835,
# 24054,
# 24484,
# 24921,
# 25030,
# 24121,
# 25195,
# 25461,
# 25503,
# 27082,
# 27087,
# 28115,
# 28518,
# 28910,
# 29052,
# 29079,
# 29136,
# 29323,
# 29522,
# 29714,
# 31195,
# 31214,
# 31661,
# 31694,
# 32183,
# 32195,
# 32304,
# 32456,
# 32519,
# 32773,
# 32786,
# 32830,
# 32841,
# 33289,
# 33643,
# 34528,
# 35359,
# 36015,
# 36857,
# 37472,
# 38106,
# 38912,
# 39104,
# 39116,
# 39163,
# 39173,
# 39968,
# 40100,
# 40220,
# 40924,
# 41643,
# 41864,
# 42681,
# 44682,
# 44791,
# 45330,
# 45393,
# 45711,
# 45978,
# 46192,
# 47452,
# 47502,
# 47589,
# 47600,
# 48078,
# 56611,
# 56662,
# 56682,
# 56776,
# 56777,
# 57235,
# 58023,
# 58032,
# 58132,
# 58424,
# 58639,
# 58657,
# 59126,
# 59261,
# 59332,
# 59345,
# 59410,
# 59774,
# 59776,
# 59779,
# 60170,
# 60309,
# 60391,
# 60592,
# 61878,
# 61947,
# 62324,
# 62409,
# 62410,
# 63001,
# 63102,
# 63233,
# 63240,
# 63415,
# 63477,
# 63674,
# 63882,
# 63901,
# 64038,
# 64056,
# 64327,
# 64400,
# 65195,
# 65986,
# 66020]).exclude(id__in=matriculados)
#     conta = 0
#     contado = 0
#     for postulante in postulantes:
#         listarequisitos = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, detalleevidenciarequisitosaspirante__estado_aprobacion=2, requisitos__requisito__claserequisito__clasificacion=1).values_list('requisitos__id', flat=True)
#         for lis in listarequisitos:
#             #COPIA A COLOR DE CERTIFICADO DE VOTACIÓN VIGENTE.
#             if lis in [970]:
#                 reqma = RequisitosMaestria.objects.get(status=True, pk=1375)
#                 if EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).exists():
#                     evi = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).order_by('-id').first()
#                     evi.requisitos = reqma
#                     evi.save()
#
#                     if EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).count() > 1:
#                         evidences = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).exclude(id=evi.id)
#                         for evid in evidences:
#                             evid.status = False
#                             evid.save()
#
#                     deta = DetalleEvidenciaRequisitosAspirante.objects.filter(status=True, evidencia=evi).order_by('-id')[0]
#                     if deta.fecha_aprobacion.date() < fvence:
#                         deta.observacion = 'EL CERTIFICADO DE VOTACIÓN NO ES EL VIGENTE'
#                         deta.observacion_aprobacion = 'EL CERTIFICADO DE VOTACIÓN NO ES EL VIGENTE'
#                         deta.estado_aprobacion = 3
#                         deta.save()
#             #CERTIFICADO DEL REGISTRO EN EL SENESCYT ( EN CASO DE SER EXTRANJERO DEBE PRESENTAR TÍTULO LEGALIZADO EN UNA EMBAJADA O CONSULADO DEL ECUADOR O CON LA APOSTILLA RESPECTIVA ).
#             if lis in [967]:
#                 reqma = RequisitosMaestria.objects.get(status=True, pk=1372)
#                 if EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).exists():
#                     evi = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).order_by('-id').first()
#                     evi.requisitos = reqma
#                     evi.save()
#
#                     if EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).count() > 1:
#                         evidences = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).exclude(id=evi.id)
#                         for evid in evidences:
#                             evid.status = False
#                             evid.save()
#             #COPIA A COLOR DE CEDULA DE CIUDADANÍA O COPIA A COLOR DEL PASAPORTE EN CASO DE SER EXTRANJERO.
#             if lis in [968]:
#                 reqma = RequisitosMaestria.objects.get(status=True, pk=1373)
#                 if EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).exists():
#                     evi = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).order_by('-id').first()
#                     evi.requisitos = reqma
#                     evi.save()
#
#                     if EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).count() > 1:
#                         evidences = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).exclude(id=evi.id)
#                         for evid in evidences:
#                             evid.status = False
#                             evid.save()
#             #CERTIFICADO DE EXPERIENCIA PROFESIONAL
#             if lis in [966]:
#                 reqma = RequisitosMaestria.objects.get(status=True, pk=1371)
#                 if EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).exists():
#                     evi = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).order_by('-id').first()
#                     evi.requisitos = reqma
#                     evi.save()
#
#                     if EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).count() > 1:
#                         evidences = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).exclude(id=evi.id)
#                         for evid in evidences:
#                             evid.status = False
#                             evid.save()
#
#             #HOJA DE VIDA.
#             if lis in [969]:
#                 reqma = RequisitosMaestria.objects.get(status=True, pk=1374)
#                 if EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).exists():
#                     evi = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).order_by('-id').first()
#                     evi.requisitos = reqma
#                     evi.save()
#
#                     if EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).count() > 1:
#                         evidences = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).exclude(id=evi.id)
#                         for evid in evidences:
#                             evid.status = False
#                             evid.save()
#
#         evidema = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, detalleevidenciarequisitosaspirante__estado_aprobacion=2, requisitos__id__in=[1375, 1372, 1374, 1371, 1373])
#         if not evidema.count() >= 5:
#             postulante.estado_aprobador = 1
#             postulante.save()
#             conta += 1
#         else:
#             if evidema.count() >= 5 and postulante.estado_aprobador == 1:
#                 postulante.estado_aprobador = 2
#                 postulante.save()
#             contado += 1
#
#         if postulante.formapagopac and postulante.formapagopac.id == 2:
#             listarequisitosfi = EvidenciaRequisitosAspirante.objects.filter(inscripcioncohorte=postulante, requisitos__requisito__claserequisito__clasificacion=3, detalleevidenciarequisitosaspirante__estado_aprobacion=2)
#             for lis in listarequisitosfi:
#                 if RequisitosMaestria.objects.filter(requisito=lis.requisitos.requisito, cohorte=cohorte2023, status=True):
#                     requi = RequisitosMaestria.objects.filter(requisito=lis.requisitos.requisito,
#                                                               cohorte_id=cohorte2023,
#                                                               status=True)[0]
#                     lis.requisitos = requi
#                     lis.save()
#
#                     if lis.requisitos.requisito.id in [54, 62]:
#                         deta = DetalleEvidenciaRequisitosAspirante.objects.filter(status=True, evidencia=lis).order_by('-id').first()
#                         if deta.fecha_aprobacion.date() < fvence:
#                             deta.estado_aprobacion = 3
#                             deta.observacion = 'EL CERTIFICADO DE VOTACIÓN NO ES EL VIGENTE'
#                             deta.observacion_aprobacion = 'EL CERTIFICADO DE VOTACIÓN NO ES EL VIGENTE'
#                             deta.save()
#
#                         if DetalleEvidenciaRequisitosAspirante.objects.filter(status=True, evidencia=lis).count() > 1:
#                             detalles = DetalleEvidenciaRequisitosAspirante.objects.filter(status=True, evidencia=lis).exclude(id=deta.id)
#                             for det in detalles:
#                                 det.status = False
#                                 det.save()
#
#         if Rubro.objects.filter(inscripcion=postulante, status=True).exists():
#             aspiranteconrubro = Rubro.objects.filter(inscripcion=postulante, status=True)
#             for crubro in aspiranteconrubro:
#                 rubro = Rubro.objects.get(status=True, id=crubro.id)
#                 rubro.status = False
#                 rubro.save()
#
#                 if rubro.idrubroepunemi != 0:
#                     cursor = connections['epunemi'].cursor()
#                     sql = """SELECT id FROM sagest_pago WHERE rubro_id=%s; """ % (rubro.idrubroepunemi)
#                     cursor.execute(sql)
#                     tienerubropagos = cursor.fetchone()
#
#                     if tienerubropagos is None:
#                         sql = """DELETE FROM sagest_rubro WHERE sagest_rubro.id=%s AND sagest_rubro.idrubrounemi=%s; """ % (rubro.idrubroepunemi, rubro.id)
#                         cursor.execute(sql)
#                         cursor.close()
#
#         postulante.cohortes_id = cohorte2023.id
#
#         if postulante.tiporespuesta:
#             if postulante.tiporespuesta.id not in [4, 2]:
#                 if postulante.status == False:
#                     postulante.status =True
#
#         postulante.tiporespuesta = None
#         postulante.save()
#
#         cont += 1
#
#         estado = 0
#         if postulante.estado_aprobador == 1:
#             estado = 'EN PROCESO'
#         elif postulante.estado_aprobador == 2:
#             estado = 'ADMITIDO'
#         elif postulante.estado_aprobador == 3:
#             estado = 'RECHAZADO'
#
#         print(cont, 'Lead:', ' ', postulante.inscripcionaspirante.persona, 'Cedula:', ' ', postulante.inscripcionaspirante.persona.cedula,'Cohorte:', ' ', postulante.cohortes, 'Estado: ', estado, 'Canti: ', postulante.total_evidencias())
#
#     # print("Migración de tiene modalidad")
#     #
#     # co = 0
#     # postulantes = InscripcionCohorte.objects.filter(status=True, fecha_creacion__year__in=[2022,2023])
#     # for postulante in postulantes:
#     #     if DetalleAprobacionFormaPago.objects.filter(status=True, inscripcion=postulante).count() > 1:
#     #         postulante.puedeeditarmp = False
#     #     else:
#     #         postulante.puedeeditarmp = True
#     #     postulante.save()
#     #     co += 1
#     #     print(co, 'Lead:', ' ', postulante, 'estado:', ' ', postulante.puedeeditarmp)
# except Exception as ex:
#     pass
#
    # idcontratos = Contrato.objects.filter(status=True).values_list('inscripcion__id', flat=True).order_by('inscripcion__id')
    # postulantes = InscripcionCohorte.objects.filter(status=True, id__in=idcontratos, aceptado=False)
    # c = 0
    # for postulante in postulantes:
    #     if postulante.tiene_contrato_subido() == 2:
    #         postulante.aceptado = True
    #         postulante.save()
    #         observacion = ''
    #         if postulante.formapagopac.id == 1:
    #             observacion = 'Aceptó modalidad de pago por contado - Proceso masivo'
    #         else:
    #             observacion = 'Aceptó modalidad de pago diferido - Proceso masivo'
    #
    #         deta = DetalleAprobacionFormaPago(inscripcion_id=postulante.id,
    #                                           formapagopac=postulante.formapagopac,
    #                                           estadoformapago=1,
    #                                           observacion=observacion,
    #                                           persona=postulante.inscripcionaspirante.persona)
    #         deta.save()
    #
    #         c += 1
    #         print(f'{c} - Postulante: {postulante.inscripcionaspirante.persona}, - Cohorte: {postulante.cohortes}')
    #
    # print("Matriculados antiguos")
    # postulantes = InscripcionCohorte.objects.filter(status=True, aceptado=False)

    # c = 0
    # for postulante in postulantes:
    #     if postulante.tiene_matricula_cohorte():
    #         postulante.aceptado = True
    #         postulante.save()
    #
    #         c += 1
    #         print(f'{c} - Postulante: {postulante.inscripcionaspirante.persona}, - Cohorte: {postulante.cohortes}')



# try:
#     print("Contratos")
#     filtro = Q(status = True, cohortes__maestriaadmision__carrera__coordinacion__id = 7, formapagopac__id = 2, estado_aprobador = 2)
#     filtro = filtro & (~Q(contrato__archivocontrato__exact='')) & (Q(contrato__estado=1))
#     postulantes = InscripcionCohorte.objects.filter(filtro).order_by('-fecha_creacion')
#     c=0
#     for postulante in postulantes:
#         contrato = Contrato.objects.get(status=True, inscripcion=postulante)
#         evi = DetalleAprobacionContrato.objects.filter(status=True, contrato_id=contrato.id, espagare=False).order_by('-id').first()
#         if contrato.estado != evi.estado_aprobacion:
#             contrato.estado = evi.estado_aprobacion
#             contrato.save()
#             c += 1
#             print(f'N°{c} - Postulante: {postulante}')
#
#     filtros = Q(status = True, cohortes__maestriaadmision__carrera__coordinacion__id = 7, formapagopac__id = 2, estado_aprobador = 2)
#     filtros = filtros & (~Q(contrato__archivopagare__exact='')) & (Q(contrato__estadopagare=1))
#     postulantes = InscripcionCohorte.objects.filter(filtros).order_by('-fecha_creacion')
#
#     print("Pagarés")
#     for postulante in postulantes:
#         contrato = Contrato.objects.get(status=True, inscripcion=postulante)
#         evi = DetalleAprobacionContrato.objects.filter(status=True, contrato_id=contrato.id, espagare=True).order_by('-id').first()
#         if contrato.estadopagare != evi.estado_aprobacion:
#             contrato.estadopagare = evi.estado_aprobacion
#             contrato.save()
#
#             print(f'Postulante: {postulante}')
#
# except Exception as ex:
#     pass
# try:
#     cohorte2023 = CohorteMaestria.objects.get(id=161, status=True)
#     cont = 0
#     matriculados = []
#     fvence = datetime.strptime('2023-02-05', '%Y-%m-%d').date()
#
#     postu = InscripcionCohorte.objects.filter(status=True, cohortes__id=143)
#     for pos in postu:
#         if pos.tiene_matricula_cohorte():
#             matriculados.append(pos.id)
#
#     postulantes = InscripcionCohorte.objects.filter(status=True, estado_aprobador__in=[1, 2], cohortes__id=143).exclude(id__in=matriculados)
#     conta = 0
#     contado = 0
#     for postulante in postulantes:
#         listarequisitos = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, detalleevidenciarequisitosaspirante__estado_aprobacion=2, requisitos__requisito__claserequisito__clasificacion=1).values_list('requisitos__id', flat=True)
#         for lis in listarequisitos:
#             #COPIA A COLOR DE CERTIFICADO DE VOTACIÓN VIGENTE.
#             if lis in [899]:
#                 reqma = RequisitosMaestria.objects.get(status=True, pk=1189)
#                 if EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).exists():
#                     evi = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).order_by('-id').first()
#                     evi.requisitos = reqma
#                     evi.save()
#
#                     if EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).count() > 1:
#                         evidences = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).exclude(id=evi.id)
#                         for evid in evidences:
#                             evid.status = False
#                             evid.save()
#
#                     deta = DetalleEvidenciaRequisitosAspirante.objects.filter(status=True, evidencia=evi).order_by('-id')[0]
#                     if deta.fecha_aprobacion.date() < fvence:
#                         deta.observacion = 'EL CERTIFICADO DE VOTACIÓN NO ES EL VIGENTE'
#                         deta.observacion_aprobacion = 'EL CERTIFICADO DE VOTACIÓN NO ES EL VIGENTE'
#                         deta.estado_aprobacion = 3
#                         deta.save()
#             #CERTIFICADO DEL REGISTRO EN EL SENESCYT ( EN CASO DE SER EXTRANJERO DEBE PRESENTAR TÍTULO LEGALIZADO EN UNA EMBAJADA O CONSULADO DEL ECUADOR O CON LA APOSTILLA RESPECTIVA ).
#             if lis in [900]:
#                 reqma = RequisitosMaestria.objects.get(status=True, pk=1190)
#                 if EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).exists():
#                     evi = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).order_by('-id').first()
#                     evi.requisitos = reqma
#                     evi.save()
#
#                     if EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).count() > 1:
#                         evidences = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).exclude(id=evi.id)
#                         for evid in evidences:
#                             evid.status = False
#                             evid.save()
#             #COPIA A COLOR DE CEDULA DE CIUDADANÍA O COPIA A COLOR DEL PASAPORTE EN CASO DE SER EXTRANJERO.
#             if lis in [898]:
#                 reqma = RequisitosMaestria.objects.get(status=True, pk=1188)
#                 if EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).exists():
#                     evi = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).order_by('-id').first()
#                     evi.requisitos = reqma
#                     evi.save()
#
#                     if EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).count() > 1:
#                         evidences = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).exclude(id=evi.id)
#                         for evid in evidences:
#                             evid.status = False
#                             evid.save()
#             #CERTIFICADO DE EXPERIENCIA DE ENSEÑANZA DE INGLÉS  MÍNIMO 1 AÑO
#             if lis in [901]:
#                 reqma = RequisitosMaestria.objects.get(status=True, pk=1202)
#                 if EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).exists():
#                     evi = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).order_by('-id').first()
#                     evi.requisitos = reqma
#                     evi.save()
#
#                     if EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).count() > 1:
#                         evidences = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).exclude(id=evi.id)
#                         for evid in evidences:
#                             evid.status = False
#                             evid.save()
#
#             #CERTIFICADO DE FORMACIÓN PEDAGÓGICA Y/O METODOLÓGICA DE INGLÉS - DURACIÓN 120 HORAS
#             if lis in [903]:
#                 reqma = RequisitosMaestria.objects.get(status=True, pk=1192)
#                 if EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).exists():
#                     evi = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).order_by('-id').first()
#                     evi.requisitos = reqma
#                     evi.save()
#
#                     if EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).count() > 1:
#                         evidences = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).exclude(id=evi.id)
#                         for evid in evidences:
#                             evid.status = False
#                             evid.save()
#             #CERTIFICADO DE NIVEL DE SUFICIENCIA DE INGLÉS B1 DE ACUERDO AL MARCO COMÚN EUROPEO DE REFERENCIAS PARA LENGUAS (MCER)
#             if lis in [902]:
#                 reqma = RequisitosMaestria.objects.get(status=True, pk=1191)
#                 if EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).exists():
#                     evi = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).order_by('-id').first()
#                     evi.requisitos = reqma
#                     evi.save()
#
#                     if EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).count() > 1:
#                         evidences = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).exclude(id=evi.id)
#                         for evid in evidences:
#                             evid.status = False
#                             evid.save()
#             #CARTA DE INTENCIÓN DE INGRESO AL PROGRAMA EN INGLÉS
#             if lis in [904]:
#                 reqma = RequisitosMaestria.objects.get(status=True, pk=1193)
#                 if EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).exists():
#                     evi = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).order_by('-id').first()
#                     evi.requisitos = reqma
#                     evi.save()
#
#                     if EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).count() > 1:
#                         evidences = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).exclude(id=evi.id)
#                         for evid in evidences:
#                             evid.status = False
#                             evid.save()
#             #CARTA DE COMPROMISO
#             if lis in [1044]:
#                 reqma = RequisitosMaestria.objects.get(status=True, pk=1220)
#                 if EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).exists():
#                     evi = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).order_by('-id').first()
#                     evi.requisitos = reqma
#                     evi.save()
#
#                     if EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).count() > 1:
#                         evidences = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).exclude(id=evi.id)
#                         for evid in evidences:
#                             evid.status = False
#                             evid.save()
#
#             #CERTIFICADO LABORAL CON EXPERIENCIA DOCENTE MÍNIMO 1 AÑO
#             # if lis in [893]:
#             #     reqma = RequisitosMaestria.objects.get(status=True, pk=1045)
#             #     evi = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).order_by('-id').first()
#             #     evi.requisitos = reqma
#             #     evi.save()
#             #
#             #     if EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).count() > 1:
#             #         evidences = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos__id=lis).exclude(id=evi.id)
#             #         for evid in evidences:
#             #             evid.status = False
#             #             evid.save()
#
#         evidema = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, detalleevidenciarequisitosaspirante__estado_aprobacion=2, requisitos__id__in=[1189, 1190, 1188, 1202, 1192, 1191, 1193, 1220])
#         if not evidema.count() >= 7:
#             postulante.estado_aprobador = 1
#             postulante.save()
#             conta += 1
#         else:
#             if evidema.count() >= 7 and postulante.estado_aprobador == 1:
#                 postulante.estado_aprobador = 2
#                 postulante.save()
#             contado += 1
#
#         if postulante.formapagopac and postulante.formapagopac.id == 2:
#             listarequisitosfi = EvidenciaRequisitosAspirante.objects.filter(inscripcioncohorte=postulante, requisitos__requisito__claserequisito__clasificacion=3, detalleevidenciarequisitosaspirante__estado_aprobacion=2)
#             for lis in listarequisitosfi:
#                 if RequisitosMaestria.objects.filter(requisito=lis.requisitos.requisito, cohorte=cohorte2023, status=True):
#                     requi = RequisitosMaestria.objects.filter(requisito=lis.requisitos.requisito,
#                                                               cohorte_id=cohorte2023,
#                                                               status=True)[0]
#                     lis.requisitos = requi
#                     lis.save()
#
#                     if lis.requisitos.requisito.id in [54, 62]:
#                         deta = DetalleEvidenciaRequisitosAspirante.objects.filter(status=True, evidencia=lis).order_by('-id').first()
#                         if deta.fecha_aprobacion.date() < fvence:
#                             deta.estado_aprobacion = 3
#                             deta.observacion = 'EL CERTIFICADO DE VOTACIÓN NO ES EL VIGENTE'
#                             deta.observacion_aprobacion = 'EL CERTIFICADO DE VOTACIÓN NO ES EL VIGENTE'
#                             deta.save()
#
#                         if DetalleEvidenciaRequisitosAspirante.objects.filter(status=True, evidencia=lis).count() > 1:
#                             detalles = DetalleEvidenciaRequisitosAspirante.objects.filter(status=True, evidencia=lis).exclude(id=deta.id)
#                             for det in detalles:
#                                 det.status = False
#                                 det.save()
#
#         if Rubro.objects.filter(inscripcion=postulante, status=True).exists():
#             aspiranteconrubro = Rubro.objects.filter(inscripcion=postulante, status=True)
#             for crubro in aspiranteconrubro:
#                 rubro = Rubro.objects.get(status=True, id=crubro.id)
#                 rubro.status = False
#                 rubro.save()
#
#                 if rubro.idrubroepunemi != 0:
#                     cursor = connections['epunemi'].cursor()
#                     sql = """SELECT id FROM sagest_pago WHERE rubro_id=%s; """ % (rubro.idrubroepunemi)
#                     cursor.execute(sql)
#                     tienerubropagos = cursor.fetchone()
#
#                     if tienerubropagos is None:
#                         sql = """DELETE FROM sagest_rubro WHERE sagest_rubro.id=%s AND sagest_rubro.idrubrounemi=%s; """ % (rubro.idrubroepunemi, rubro.id)
#                         cursor.execute(sql)
#                         cursor.close()
#
#         postulante.cohortes_id = cohorte2023.id
#         # if postulante.tiporespuesta:
#         #     postulante.tiporespuesta = 1
#         postulante.save()
#         cont += 1
#
#         estado = 0
#         if postulante.estado_aprobador == 1:
#             estado = 'EN PROCESO'
#         elif postulante.estado_aprobador == 2:
#             estado = 'ADMITIDO'
#         elif postulante.estado_aprobador == 3:
#             estado = 'RECHAZADO'
#
#         print(cont, 'Lead:', ' ', postulante.inscripcionaspirante.persona, 'Cedula:', ' ', postulante.inscripcionaspirante.persona.cedula,'Cohorte:', ' ', postulante.cohortes, 'Estado: ', estado, 'Canti: ', postulante.total_evidencias())
# except Exception as ex:
#     print('error: %s' % ex)
# def traer_datos_query():
#     try:
#         sql="""
#                 SELECT coordinacion.nombre AS facultad, carrera.nombre AS carrera,
#                 (persona.apellido1||' '||persona.apellido2||' '||persona.nombres) AS estudiante,
#                 sesion.nombre AS seccion,
#                 (ARRAY_TO_STRING(array(SELECT DISTINCT sesion0.nombre FROM sga_materia materia0
#                 INNER JOIN sga_nivel nivel0 ON materia0.nivel_id = nivel0.id
#                 INNER JOIN sga_sesion sesion0 ON nivel0.sesion_id = sesion0.id
#                 INNER JOIN sga_asignaturamalla asignaturamalla0 ON materia0.asignaturamalla_id = asignaturamalla0.id
#                 INNER JOIN sga_malla malla0 ON asignaturamalla0.malla_id = malla0.id
#                 WHERE malla0.carrera_id=carrera.id AND nivel0.periodo_id=224
#                 AND materia0."status" ORDER BY sesion0.nombre), ', ')) AS sesiones,
#                 inscripcion.id AS inscripcion_id,
#                 sesion.id AS seccion_id,
#                 (ARRAY_TO_STRING(array(SELECT DISTINCT sesion0.id FROM sga_materia materia0
#                 INNER JOIN sga_nivel nivel0 ON materia0.nivel_id = nivel0.id
#                 INNER JOIN sga_sesion sesion0 ON nivel0.sesion_id = sesion0.id
#                 INNER JOIN sga_asignaturamalla asignaturamalla0 ON materia0.asignaturamalla_id = asignaturamalla0.id
#                 INNER JOIN sga_malla malla0 ON asignaturamalla0.malla_id = malla0.id
#                 WHERE malla0.carrera_id=carrera.id AND nivel0.periodo_id=224
#                 AND materia0."status" ORDER BY sesion0.id), ', ')) AS sesiones_id
#                 FROM sga_matricula matricula
#                 INNER JOIN sga_inscripcion inscripcion ON matricula.inscripcion_id = inscripcion.id
#                 INNER JOIN sga_nivel nivel ON matricula.nivel_id = nivel.id
#                 INNER JOIN sga_carrera carrera ON inscripcion.carrera_id = carrera.id
#                 INNER JOIN sga_coordinacion_carrera corcar ON corcar.carrera_id= carrera.id
#                 INNER JOIN sga_coordinacion coordinacion ON coordinacion.id = corcar.coordinacion_id
#                 INNER JOIN sga_sesion sesion ON inscripcion.sesion_id = sesion.id
#                 INNER JOIN sga_persona persona ON inscripcion.persona_id = persona.id
#                 WHERE nivel.periodo_id = 177
#                 AND inscripcion."status"
#                 AND matricula.retiradomatricula = FALSE
#                 AND matricula."status"
#                 AND coordinacion.id IN (1, 2, 3, 4, 5)
#                 ORDER BY coordinacion.nombre, carrera.nombre
#         """
#         cursor = connections['default'].cursor()
#         cursor.execute(sql)
#         rows_effected = cursor.rowcount
#         listado = cursor.fetchall()
#         campos = [desc[0] for desc in cursor.description]
#         directory = os.path.join(MEDIA_ROOT, 'reportes', 'gestion')
#         try:
#             os.stat(directory)
#         except:
#             os.mkdir(directory)
#         name_document = 'reporte'
#         nombre_archivo = name_document + "_1.xlsx"
#         directory = os.path.join(MEDIA_ROOT, 'reportes', 'gestion', nombre_archivo)
#
#         __author__ = 'Unemi'
#         workbook = xlsxwriter.Workbook(directory, {'constant_memory': True})
#         ws = workbook.add_worksheet('resultados')
#         fuentecabecera = workbook.add_format({
#             'align': 'center',
#             'bg_color': 'silver',
#             'border': 1,
#             'bold': 1
#         })
#
#         formatoceldacenter = workbook.add_format({
#             'border': 1,
#             'valign': 'vcenter',
#             'align': 'center'})
#
#         row_num, numcolum = 0, 0
#
#         for col_num in campos:
#             ws.write(row_num, numcolum, col_num, fuentecabecera)
#             ws.set_column(row_num, numcolum, 40)
#             numcolum += 1
#         row_num += 1
#         for lis in listado:
#             colum_num = 0
#             for l in lis:
#                 ws.write(row_num, colum_num, l, formatoceldacenter)
#                 ws.set_column(row_num, numcolum, 40)
#                 colum_num += 1
#             row_num += 1
#
#         workbook.close()
#         response = HttpResponse(directory,
#                                 content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#         response['Content-Disposition'] = 'attachment; filename=%s' % name_document
#         #
#         url_file = "{}reportes/gestion/{}".format(MEDIA_URL, nombre_archivo)
#         print(url_file)
#     except Exception as ex:
#         pass
#
# traer_datos_query()
#
# print("Finaliza")

# print(u"********************POSTULANTES CON EVIDENCIAS COMPLETAS GENERAL***************************")
#
# try:
#     c = 0
#     postulantes = InscripcionCohorte.objects.filter(status=True, todosubido=False)
#     for postulante in postulantes:
#         requistosmaestria = RequisitosMaestria.objects.filter(status=True, cohorte=postulante.cohortes,
#                                                               requisito__claserequisito__clasificacion__id=1,
#                                                               obligatorio=True).values_list('id', flat=True)
#         cont = 0
#         estado = False
#         for requisto in requistosmaestria:
#             if EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante,
#                                                            requisitos_id=requisto,
#                                                            requisitos__requisito__claserequisito__clasificacion__id=1).exists():
#                 evi = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante,
#                                                                   requisitos_id=requisto,
#                                                                   requisitos__requisito__claserequisito__clasificacion__id=1).order_by(
#                     '-id').first()
#                 deta = DetalleEvidenciaRequisitosAspirante.objects.filter(status=True, evidencia=evi).order_by(
#                     '-id').first()
#                 if deta.estadorevision == 1 or deta.estado_aprobacion == 2:
#                     cont += 1
#
#         if cont == requistosmaestria.count():
#             postulante.todosubido = True
#             postulante.save()
#             c += 1
#
#             print(
#                 f"N: {c} - Postulante: {postulante.inscripcionaspirante.persona} - Cohorte: {postulante.cohortes} - Cantidad: {cont} - Asesor: {postulante.asesor.persona if postulante.asesor else'No registra'}")
#
# except Exception as ex:
#     print(ex)

#     c = 0
#     lista_postu = []
#     postulantes = InscripcionCohorte.objects.filter(status=True, estado_aprobador=1, todosubido=True)
#     for postulante in postulantes:
#         requistosmaestria = RequisitosMaestria.objects.filter(status=True, cohorte=postulante.cohortes, requisito__claserequisito__clasificacion__id=1, obligatorio=True).values_list('id', flat=True)
#         estado = False
#         for requisto in requistosmaestria:
#             if EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos_id=requisto, requisitos__requisito__claserequisito__clasificacion__id=1).exists():
#                 evi = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, requisitos_id=requisto, requisitos__requisito__claserequisito__clasificacion__id=1).order_by('-id').first()
#                 deta = DetalleEvidenciaRequisitosAspirante.objects.filter(status=True, evidencia=evi).order_by('-id').first()
#                 if deta.estado_aprobacion == 3:
#                     lista_postu.append(postulante.id)
#                     break
#
#     postulante_rec = InscripcionCohorte.objects.filter(status=True, id__in=lista_postu)
#     for po in postulante_rec:
#         po.tienerechazo = True
#         po.todosubido = False
#         po.preaprobado=False
#         po.save()
#         c += 1
#         print(f"N: {c} - Postulante: {po.inscripcionaspirante.persona} - Cohorte: {po.cohortes} - Asesor: {po.asesor.persona}")
#
# except Exception as ex:
#     print(ex)



# traer_datos_query()
# try:
#     # postulantes = InscripcionCohorte.objects.filter(status=True, asesor__id__in=[24, 25])
#     # listado_pagos = []
#     # for postulante in postulantes:
#     #     listado_fechas = []
#     #     if postulante.comprobante_subido():
#     #         listado_fechas.append(postulante.get_comprobante_subido().fecha_creacion.date())
#     #     if postulante.comprobante_subido_epunemi():
#     #         listado_fechas.append(postulante.get_comprobante_subido_epunemi())
#     #     if postulante.total_pagado_rubro_cohorte() > 0:
#     #         listado_fechas.append(postulante.fecha_primer_pago())
#     #
#     #     if len(listado_fechas) > 0:
#     #         listado_pagos.append(postulante.id)
#     #
#     # leads = InscripcionCohorte.objects.filter(status=True, asesor__id__in=[24, 25]).exclude(id__in=listado_pagos)
#     # asesores = [23, 20, 22, 21, 6]
#     # c = a = 0
#     # observacion = 'ASIGNACIÓN MASIVA'
#     # for lead in leads:
#     #     if AsesorComercial.objects.filter(status=True, pk=int(asesores[c])).exists():
#     #         asesor = AsesorComercial.objects.get(status=True, pk=int(asesores[c]))
#     #         asesoranti = lead.asesor
#     #         lead.asesor = asesor
#     #         lead.estado_asesor = 2
#     #         lead.tiporespuesta_id = None
#     #         lead.save()
#     #         if not asesoranti:
#     #             histo = HistorialAsesor(inscripcion=lead, fecha_inicio=lead.fecha_modificacion,
#     #                             fecha_fin=None, asesor=lead.asesor, observacion=observacion)
#     #             histo.save()
#     #         else:
#     #             histoanti = HistorialAsesor.objects.get(inscripcion_id=lead.id, fecha_fin=None)
#     #             histoanti.fecha_fin = lead.fecha_modificacion
#     #             histoanti.save()
#     #             histo = HistorialAsesor(inscripcion_id=lead.id, fecha_inicio=lead.fecha_modificacion,
#     #                                     fecha_fin=None, asesor=lead.asesor, observacion=observacion)
#     #             histo.save()
#     #
#     #     a += 1
#     #     if c == 4:
#     #         c = 0
#     #     else:
#     #         c += 1
#     #
#     #     print(f"{a} - Lead: {lead.inscripcionaspirante.persona} - Cohorte: {lead.cohortes} - Asesor: {lead.asesor.persona}")

# cohorte2023 = CohorteMaestria.objects.get(id=148, status=True)
    # cont = 0
    # matriculados = []
    #
    # postu = InscripcionCohorte.objects.filter(status=True, cohortes__id=141)
    # for pos in postu:
    #     if pos.tiene_matricula_cohorte():
    #         matriculados.append(pos.id)
    #
    # postulantes = InscripcionCohorte.objects.filter(status=True, cohortes__id=141).exclude(id__in=matriculados)
    # conta = 0
    # contado = 0
    # for postulante in postulantes:
    #     listarequisitos = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, detalleevidenciarequisitosaspirante__estado_aprobacion=2, requisitos__requisito__claserequisito__clasificacion=1).values_list('requisitos__id', flat=True)
    #     for lis in listarequisitos:
    #         #COPIA A COLOR CÉDULA DE CIUDADANÍA O COPIA A COLOR DEL PASAPORTE EN CASO DE SER EXTRANJERO.
    #         if lis in [879, 355, 412]:
    #             reqma = RequisitosMaestria.objects.get(status=True, pk=1003)
    #             evi = EvidenciaRequisitosAspirante.objects.get(status=True, inscripcioncohorte=postulante, requisitos__id=lis)
    #             evi.requisitos = reqma
    #             evi.save()
    #             if lis in [355, 412]:
    #                 deta = DetalleEvidenciaRequisitosAspirante.objects.filter(status=True, evidencia=evi).order_by('-id')[0]
    #                 deta.observacion = ''
    #                 deta.observacion_aprobacion = ''
    #                 deta.estado_aprobacion = 1
    #                 deta.save()
    #         #COPIA A COLOR DE CERTIFICADO DE VOTACIÓN VIGENTE.
    #         if lis in [877]:
    #             reqma = RequisitosMaestria.objects.get(status=True, pk=1001)
    #             evi = EvidenciaRequisitosAspirante.objects.get(status=True, inscripcioncohorte=postulante, requisitos__id=lis)
    #             evi.requisitos = reqma
    #             evi.save()
    #             deta = DetalleEvidenciaRequisitosAspirante.objects.filter(status=True, evidencia=evi).order_by('-id')[0]
    #             deta.observacion = 'EL CERTIFICADO DE VOTACIÓN NO ES EL VIGENTE'
    #             deta.observacion_aprobacion = 'EL CERTIFICADO DE VOTACIÓN NO ES EL VIGENTE'
    #             deta.estado_aprobacion = 3
    #             deta.save()
    #         #CERTIFICADO DEL REGISTRO EN EL SENESCYT ( EN CASO DE SER EXTRANJERO DEBE PRESENTAR TÍTULO LEGALIZADO EN UNA EMBAJADA O CONSULADO DEL ECUADOR O CON LA APOSTILLA RESPECTIVA ).
    #         if lis in [878, 358, 415]:
    #             reqma = RequisitosMaestria.objects.get(status=True, pk=1002)
    #             evi = EvidenciaRequisitosAspirante.objects.get(status=True, inscripcioncohorte=postulante, requisitos__id=lis)
    #             evi.requisitos = reqma
    #             evi.save()
    #             if lis in [358, 415]:
    #                 deta = DetalleEvidenciaRequisitosAspirante.objects.filter(status=True, evidencia=evi).order_by('-id')[0]
    #                 deta.observacion = ''
    #                 deta.observacion_aprobacion = ''
    #                 deta.estado_aprobacion = 1
    #                 deta.save()
    #
    #         #CERTIFICADO LABORAL CON EXPERIENCIA DOCENTE MÍNIMO 1 AÑO
    #         if lis in [882, 398, 416]:
    #             reqma = RequisitosMaestria.objects.get(status=True, pk=1004)
    #             evi = EvidenciaRequisitosAspirante.objects.get(status=True, inscripcioncohorte=postulante, requisitos__id=lis)
    #             evi.requisitos = reqma
    #             evi.save()
    #             if lis in [398, 416]:
    #                 deta = DetalleEvidenciaRequisitosAspirante.objects.filter(status=True, evidencia=evi).order_by('-id')[0]
    #                 deta.observacion = ''
    #                 deta.observacion_aprobacion = ''
    #                 deta.estado_aprobacion = 1
    #                 deta.save()
    #
    #     evidema = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, detalleevidenciarequisitosaspirante__estado_aprobacion=2, requisitos__id__in=[1003, 1001, 1002, 1004])
    #     if not evidema.count() == 4:
    #         postulante.estado_aprobador = 1
    #         postulante.save()
    #         conta += 1
    #     else:
    #         if evidema.count() == 4 and postulante.estado_aprobador == 1:
    #             postulante.estado_aprobador = 2
    #             postulante.save()
    #         contado += 1
    #
    #     if postulante.formapagopac and postulante.formapagopac.id == 2:
    #         listarequisitosfi = EvidenciaRequisitosAspirante.objects.filter(inscripcioncohorte=postulante, requisitos__requisito__claserequisito__clasificacion=3, detalleevidenciarequisitosaspirante__estado_aprobacion=2)
    #         for lis in listarequisitosfi:
    #             if RequisitosMaestria.objects.filter(requisito=lis.requisitos.requisito, cohorte=cohorte2023, status=True):
    #                 requi = RequisitosMaestria.objects.filter(requisito=lis.requisitos.requisito,
    #                                                           cohorte_id=cohorte2023,
    #                                                           status=True)[0]
    #                 lis.requisitos = requi
    #                 lis.save()
    #
    #                 if lis.requisitos.requisito.id in [54, 62]:
    #                     deta = DetalleEvidenciaRequisitosAspirante.objects.filter(status=True, evidencia=lis).order_by('-id').first()
    #                     deta.estado_aprobacion = 3
    #                     deta.observacion = 'EL CERTIFICADO DE VOTACIÓN NO ES EL VIGENTE'
    #                     deta.observacion_aprobacion = 'EL CERTIFICADO DE VOTACIÓN NO ES EL VIGENTE'
    #                     deta.save()
    #
    #                     if DetalleEvidenciaRequisitosAspirante.objects.filter(status=True, evidencia=lis).count() > 1:
    #                         detalles = DetalleEvidenciaRequisitosAspirante.objects.filter(status=True, evidencia=lis).exclude(id=deta.id)
    #                         for det in detalles:
    #                             det.status = False
    #                             det.save()
    #
    #     if Rubro.objects.filter(inscripcion=postulante, status=True).exists():
    #         aspiranteconrubro = Rubro.objects.filter(inscripcion=postulante, status=True)
    #         for crubro in aspiranteconrubro:
    #             rubro = Rubro.objects.get(status=True, id=crubro.id)
    #             rubro.status = False
    #             rubro.save()
    #
    #             if rubro.idrubroepunemi != 0:
    #                 cursor = connections['epunemi'].cursor()
    #                 sql = """SELECT id FROM sagest_pago WHERE rubro_id=%s; """ % (rubro.idrubroepunemi)
    #                 cursor.execute(sql)
    #                 tienerubropagos = cursor.fetchone()
    #
    #                 if tienerubropagos is None:
    #                     sql = """DELETE FROM sagest_rubro WHERE sagest_rubro.id=%s AND sagest_rubro.idrubrounemi=%s; """ % (rubro.idrubroepunemi, rubro.id)
    #                     cursor.execute(sql)
    #                     cursor.close()
    #
    #     postulante.cohortes_id = cohorte2023.id
    #     # if postulante.tiporespuesta:
    #     #     postulante.tiporespuesta = 1
    #     postulante.save()
    #     cont += 1
    #
    #     estado = 0
    #     if postulante.estado_aprobador == 1:
    #         estado = 'EN PROCESO'
    #     elif postulante.estado_aprobador == 2:
    #         estado = 'ADMITIDO'
    #     elif postulante.estado_aprobador == 3:
    #         estado = 'RECHAZADO'
    #
    #     print(cont, 'Lead:', ' ', postulante.inscripcionaspirante.persona, 'Cedula:', ' ', postulante.inscripcionaspirante.persona.cedula,'Cohorte:', ' ', postulante.cohortes, 'Estado: ', estado, 'FP: ', postulante.formapagopac.descripcion)
# except Exception as ex:
#         print('error: %s' % ex)
#
# print("Finaliza")


# miarchivo = openpyxl.load_workbook("Listado_de_postulantes_atendidos_asesores_nuevo.xlsx")
    # lista = miarchivo.get_sheet_by_name('resultados')
    # totallista = lista.rows
    # a=0
    # c=0
    # for filas in totallista:
    #     a += 1
    #     if a > 1:
    #         idinscripcioncohorte = int(filas[0].value)
    #         cedula = str((filas[9].value)).replace(".", "")
    #         observacion = 'MIGRAXIÓN MASIVA DE LEADS ATENDIDOS DESDE EL USUARIO DE AMY TORRES'
    #         if idinscripcioncohorte != 'None':
    #             if AsesorComercial.objects.filter(status=True, persona__cedula=cedula).exists():
    #                 asesor = AsesorComercial.objects.get(status=True, persona__cedula=cedula)
    #                 if InscripcionCohorte.objects.filter(status=True, id=idinscripcioncohorte).exists():
    #                     inscrito = InscripcionCohorte.objects.get(status=True, id=idinscripcioncohorte)
    #                     asesoranti = inscrito.asesor
    #                     inscrito.asesor = asesor
    #                     inscrito.estado_asesor = 2
    #                     inscrito.save()
    #                     if not asesoranti:
    #                         histo = HistorialAsesor(inscripcion=inscrito, fecha_inicio=inscrito.fecha_modificacion,
    #                                         fecha_fin=None, asesor=inscrito.asesor, observacion=observacion)
    #                         histo.save()
    #                     else:
    #                         if asesoranti.id != inscrito.asesor.id:
    #                             histoanti = HistorialAsesor.objects.get(inscripcion=inscrito, fecha_fin=None)
    #                             histoanti.fecha_fin = inscrito.fecha_modificacion
    #                             histoanti.save()
    #                             histo = HistorialAsesor(inscripcion=inscrito, fecha_inicio=inscrito.fecha_modificacion,
    #                                             fecha_fin=None, asesor=inscrito.asesor, observacion=observacion)
    #                             histo.save()
    #                         print(f"{c} - Lead: {inscrito.inscripcionaspirante.persona} - Cohorte: {inscrito.cohortes} - Asesor: {inscrito.asesor.persona}")
    #                 c += 1
    #             else:
    #                 print(u"Este asesor no existe")
    #
    #         else:
    #             print(u"FINALIZADO")
    #             break


# try:
#     cohorte2023 = CohorteMaestria.objects.get(id=146, status=True)
#     cont = 0
#     postulantes = InscripcionCohorte.objects.filter(status=True, cohortes__id=132).exclude(id__in=[32878, 35002, 34618, 32954, 32930,
#     36003, 33010, 31422, 33446, 36732, 31371, 32550, 32756, 31831, 33876, 31245, 33414, 32540, 35297, 34569, 32090, 32940, 35463, 33286, 31907,
#     33862, 33380, 25451, 32719, 32762, 34370, 34213, 31911, 34840, 31343, 34885, 34141, 32236, 34662, 34514, 35827, 33056, 28499, 35704, 34102,
#     33285, 35333, 29922, 31363, 32948, 34816, 29378, 34892, 22703, 31992, 33265, 35226, 33907, 34753, 34336, 34179, 33125, 33908, 36037, 34841,
#     35323, 34409, 32931, 34591, 34654, 31367, 34037, 31395, 35152, 31518, 36468, 34344, 25715, 27667, 34294, 33011, 32721, 34387, 32706, 33984,
#     34614, 32553, 29461, 34107, 32695, 35165, 36557, 36384, 37240, 32693, 38226, 37875, 37576, 36807, 37280, 37782, 35520, 33887, 36944, 37610,
#     35770, 37600, 38874, 31530, 31533, 37433, 35356, 34534, 39170, 37241, 39024, 38806, 39044, 38505, 37869, 38722, 39432, 38297, 39043, 37184,
#     39144, 22685, 39642, 35362, 39976, 35495, 38293, 35788, 37608, 37248, 35364, 38152, 38153, 38530, 39752, 37894, 40245, 40148, 40262, 38996,
#     40467, 39379, 37502, 39581, 25209, 36546, 37868, 37344, 39529, 40284, 38010, 24868, 39935, 36808, 23746, 39578, 40533, 39377, 40594, 22241,
#     40788, 40890, 40110, 34671, 36551, 28990, 31175, 34540, 33186, 31767, 35891, 38525, 35379, 32802, 33123, 23142, 32812, 34597, 34743, 36245,
#     29569, 36343, 36259, 39771, 41300, 29675, 40837, 40217, 39260, 40912, 40696, 41092, 35983, 22217, 23083, 38465, 38110, 40636, 38285, 38934,
#     39182, 39776, 41120, 38489, 25555, 23736, 41665, 41615, 41713, 41355, 38314, 40911, 31969, 35582, 41566, 33474, 33352, 26906, 40969, 41736,
#     34232, 40817, 41762, 40637, 34529, 37416, 41504, 40255, 42665, 42675, 42839, 41476, 43549, 26010])
#     conta = 0
#     contado = 0
#     for postulante in postulantes:
#         listarequisitos = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, detalleevidenciarequisitosaspirante__estado_aprobacion=2, requisitos__requisito__claserequisito__clasificacion=1).values_list('requisitos__id', flat=True)
#         for lis in listarequisitos:
#             #COPIA A COLOR DE CEDULA DE CIUDADANÍA O COPIA A COLOR DEL PASAPORTE EN CASO DE SER EXTRANJERO.
#             if lis in [550, 732]:
#                 reqma = RequisitosMaestria.objects.get(status=True, pk=968)
#                 evi = EvidenciaRequisitosAspirante.objects.get(status=True, inscripcioncohorte=postulante, requisitos__id=lis)
#                 evi.requisitos = reqma
#                 evi.save()
#             #COPIA A COLOR DE CERTIFICADO DE VOTACIÓN VIGENTE.
#             if lis in [551, 735]:
#                 reqma = RequisitosMaestria.objects.get(status=True, pk=970)
#                 evi = EvidenciaRequisitosAspirante.objects.get(status=True, inscripcioncohorte=postulante, requisitos__id=lis)
#                 evi.requisitos = reqma
#                 evi.save()
#                 deta = DetalleEvidenciaRequisitosAspirante.objects.filter(status=True, evidencia=evi).order_by('-id')[0]
#                 deta.observacion = 'EL CERTIFICADO DE VOTACIÓN NO ES EL VIGENTE'
#                 deta.observacion_aprobacion = 'EL CERTIFICADO DE VOTACIÓN NO ES EL VIGENTE'
#                 deta.estado_aprobacion = 3
#                 deta.save()
#             #HOJA DE VIDA.
#             if lis in [552, 733]:
#                 reqma = RequisitosMaestria.objects.get(status=True, pk=969)
#                 evi = EvidenciaRequisitosAspirante.objects.get(status=True, inscripcioncohorte=postulante, requisitos__id=lis)
#                 evi.requisitos = reqma
#                 evi.save()
#             #CERTIFICADO DEL REGISTRO EN EL SENESCYT ( EN CASO DE SER EXTRANJERO DEBE PRESENTAR TÍTULO LEGALIZADO EN UNA EMBAJADA O CONSULADO DEL ECUADOR O CON LA APOSTILLA RESPECTIVA ).
#             if lis in [553, 747]:
#                 reqma = RequisitosMaestria.objects.get(status=True, pk=967)
#                 evi = EvidenciaRequisitosAspirante.objects.get(status=True, inscripcioncohorte=postulante, requisitos__id=lis)
#                 evi.requisitos = reqma
#                 evi.save()
#             #CERTIFICADO LABORAL MINIMO 1 AÑO DE EXPERIENCIA.
#             if lis in [569, 746]:
#                 reqma = RequisitosMaestria.objects.get(status=True, pk=966)
#                 evi = EvidenciaRequisitosAspirante.objects.get(status=True, inscripcioncohorte=postulante, requisitos__id=lis)
#                 evi.requisitos = reqma
#                 evi.save()
#         evidema = EvidenciaRequisitosAspirante.objects.filter(status=True, inscripcioncohorte=postulante, detalleevidenciarequisitosaspirante__estado_aprobacion=2, requisitos__id__in=[966, 967, 968, 969, 970])
#         if not evidema.count() == 4:
#             postulante.estado_aprobador = 1
#             postulante.save()
#             conta += 1
#         else:
#             if evidema.count() == 4 and postulante.estado_aprobador == 1:
#                 postulante.estado_aprobador = 2
#                 postulante.save()
#             contado += 1
#
#         if postulante.formapagopac and postulante.formapagopac.id == 2:
#             listarequisitosfi = EvidenciaRequisitosAspirante.objects.filter(inscripcioncohorte=postulante, requisitos__requisito__claserequisito__clasificacion=3, detalleevidenciarequisitosaspirante__estado_aprobacion=2)
#             for lis in listarequisitosfi:
#                 if RequisitosMaestria.objects.filter(requisito=lis.requisitos.requisito, cohorte=cohorte2023, status=True):
#                     requi = RequisitosMaestria.objects.filter(requisito=lis.requisitos.requisito,
#                                                               cohorte_id=cohorte2023,
#                                                               status=True)[0]
#                     lis.requisitos = requi
#                     lis.save()
#
#                     if lis.requisitos.requisito.id in [54, 62]:
#                         deta = DetalleEvidenciaRequisitosAspirante.objects.get(status=True, evidencia=lis)
#                         deta.estado_aprobacion = 3
#                         deta.observacion = 'EL CERTIFICADO DE VOTACIÓN NO ES EL VIGENTE'
#                         deta.observacion_aprobacion = 'EL CERTIFICADO DE VOTACIÓN NO ES EL VIGENTE'
#                         deta.save()
#
#         if Rubro.objects.filter(inscripcion=postulante, status=True).exists():
#             aspiranteconrubro = Rubro.objects.filter(inscripcion=postulante, status=True)
#             for crubro in aspiranteconrubro:
#                 rubro = Rubro.objects.get(status=True, id=crubro.id)
#                 rubro.status = False
#                 rubro.save()
#
#                 if rubro.idrubroepunemi != 0:
#                     cursor = connections['epunemi'].cursor()
#                     sql = """DELETE FROM sagest_rubro WHERE sagest_rubro.id=%s AND sagest_rubro.idrubrounemi=%s; """ % (rubro.idrubroepunemi, rubro.id)
#                     cursor.execute(sql)
#                     cursor.close()
#
#         postulante.cohortes_id = cohorte2023.id
#         postulante.save()
#         cont += 1
#
#         estado = 0
#         if postulante.estado_aprobador == 1:
#             estado = 'EN PROCESO'
#         elif postulante.estado_aprobador == 2:
#             estado = 'ADMITIDO'
#         elif postulante.estado_aprobador == 3:
#             estado = 'RECHAZADO'
#
#         print(cont, 'Lead:', ' ', postulante.inscripcionaspirante.persona, 'Cedula:', ' ', postulante.inscripcionaspirante.persona.cedula,'Cohorte:', ' ', postulante.cohortes, 'Estado: ', estado)
#
# except Exception as ex:
#     print('error: %s' % ex)
#
# print("Finaliza")



    # aspirantesmatri = []
    # aspirantes = InscripcionCohorte.objects.filter(status=True, cohortes__id__in=[91, 100])
    #
    # for aspi in aspirantes:
    #     if aspi.estado_aprobador == 3:
    #         aspirantesmatri.append(aspi.id)
    #     elif aspi.inscripcion:
    #         if Matricula.objects.filter(status=True, inscripcion__id=aspi.inscripcion.id).exists():
    #             aspirantesmatri.append(aspi.id)
     #
    # aspirantes = InscripcionCohorte.objects.filter(status=True, cohortes__id__in=[91, 100]).exclude(id__in=aspirantesmatri)
    #
    # cont = 0
    # cohorte2023 = CohorteMaestria.objects.get(id=141, status=True)
    # for aspirante in aspirantes:
    #     listarequisitos = EvidenciaRequisitosAspirante.objects.filter(inscripcioncohorte=aspirante)
    #     listarequisitosfinan = EvidenciaRequisitosAspirante.objects.filter(inscripcioncohorte=aspirante,
    #                                                                        requisitos__requisito__claserequisito__clasificacion=3)
    #     for lis in listarequisitos:
    #         if RequisitosMaestria.objects.filter(requisito=lis.requisitos.requisito,
    #                                              cohorte_id=cohorte2023.id, status=True):
    #             requi = RequisitosMaestria.objects.filter(requisito=lis.requisitos.requisito,
    #                                                       cohorte_id=cohorte2023.id,
    #                                                       status=True)[0]
    #             lis.requisitos = requi
    #             lis.save()
    #
    #     if aspirante.formapagopac:
    #         if aspirante.formapagopac.id == 2:
    #             for listf in listarequisitosfinan:
    #                 if RequisitosMaestria.objects.filter(requisito=listf.requisitos.requisito,
    #                                                      cohorte_id=cohorte2023.id, status=True,
    #                                                      requisito__claserequisito__clasificacion=3):
    #                     requifinan = RequisitosMaestria.objects.filter(requisito=listf.requisitos.requisito,
    #                                                                    cohorte_id=cohorte2023.id, status=True,
    #                                                                    requisito__claserequisito__clasificacion=3)[0]
    #                     listf.requisitos = requifinan
    #                     listf.save()
    #
    #     if Rubro.objects.filter(inscripcion=aspirante, status=True).exists():
    #         rubros = Rubro.objects.filter(inscripcion=aspirante, status=True)
    #         for rubro in rubros:
    #             if rubro.admisionposgradotipo == 2:
    #                 qs_anterior = list(Rubro.objects.filter(pk=rubro.id).values())
    #                 tiporubroarancel = TipoOtroRubro.objects.get(pk=2845)
    #                 rubro.nombre = tiporubroarancel.nombre + ' - ' + cohorte2023.maestriaadmision.descripcion + ' - ' + cohorte2023.descripcion
    #                 rubro.cohortemaestria = cohorte2023
    #                 rubro.save()
    #                 # GUARDA AUDITORIA RUBRO
    #                 # qs_nuevo = list(Rubro.objects.filter(pk=rubro.id).values())
    #                 # GUARDA AUDITORIA RUBRO
    #                 if rubro.epunemi and rubro.idrubroepunemi > 0:
    #                     cursor = connections['epunemi'].cursor()
    #                     sql = """UPDATE sagest_rubro SET nombre='%s' WHERE sagest_rubro.status=true and sagest_rubro.id='%s'; """ % (rubro.nombre, rubro.idrubroepunemi)
    #                     # sql = "UPDATE sagest_rubro SET status=false WHERE sagest_rubro.status=true and sagest_rubro.id=" + str(rubro.idrubroepunemi)
    #                     cursor.execute(sql)
    #                     cursor.close()
    #
    #             elif rubro.admisionposgradotipo == 3:
    #                 qs_anterior = list(Rubro.objects.filter(pk=rubro.id).values())
    #                 rubro.nombre = cohorte2023.maestriaadmision.descripcion + ' - ' + cohorte2023.descripcion
    #                 rubro.cohortemaestria = cohorte2023
    #                 rubro.save()
    #                 # GUARDA AUDITORIA RUBRO
    #                 # qs_nuevo = list(Rubro.objects.filter(pk=rubro.id).values())
    #                 # GUARDA AUDITORIA RUBRO
    #                 if rubro.epunemi and rubro.idrubroepunemi > 0:
    #                     cursor = connections['epunemi'].cursor()
    #                     sql = """UPDATE sagest_rubro SET nombre='%s' WHERE sagest_rubro.status=true and sagest_rubro.id='%s'; """ % (rubro.nombre, rubro.idrubroepunemi)
    #                     cursor.execute(sql)
    #                     cursor.close()
    #
    #     aspirante.cohortes_id = cohorte2023.id
    #     aspirante.save()
    #
    #     cont += 1
    #     print(cont, 'Lead:', ' ', aspirante.inscripcionaspirante.persona, 'Cedula:', ' ', aspirante.inscripcionaspirante.persona.cedula,'Cohorte:', ' ', aspirante.cohortes)

    # print(u"********************Biotecnología***************************")
    #
    # aspirantesmatri = []
    # aspirantes = InscripcionCohorte.objects.filter(status=True, cohortes__id=107)
    #
    # for aspi in aspirantes:
    #     if aspi.inscripcion:
    #         if Matricula.objects.filter(status=True, inscripcion__id=aspi.inscripcion.id).exists():
    #             aspirantesmatri.append(aspi.id)
    #
    # aspirantes = InscripcionCohorte.objects.filter(status=True, cohortes__id=107).exclude(id__in=aspirantesmatri)
    #
    # cont = 0
    # cohorte2023 = CohorteMaestria.objects.get(id=139, status=True)
    # for aspirante in aspirantes:
    #     listarequisitos = EvidenciaRequisitosAspirante.objects.filter(inscripcioncohorte=aspirante)
    #     listarequisitosfinan = EvidenciaRequisitosAspirante.objects.filter(inscripcioncohorte=aspirante,
    #                                                                        requisitos__requisito__claserequisito__clasificacion=3)
    #     for lis in listarequisitos:
    #         if RequisitosMaestria.objects.filter(requisito=lis.requisitos.requisito,
    #                                              cohorte_id=cohorte2023.id, status=True):
    #             requi = RequisitosMaestria.objects.filter(requisito=lis.requisitos.requisito,
    #                                                       cohorte_id=cohorte2023.id,
    #                                                       status=True)[0]
    #             lis.requisitos = requi
    #             lis.save()
    #
    #     if aspirante.formapagopac:
    #         if aspirante.formapagopac.id == 2:
    #             for listf in listarequisitosfinan:
    #                 if RequisitosMaestria.objects.filter(requisito=listf.requisitos.requisito,
    #                                                      cohorte_id=cohorte2023.id, status=True,
    #                                                      requisito__claserequisito__clasificacion=3):
    #                     requifinan = RequisitosMaestria.objects.filter(requisito=listf.requisitos.requisito,
    #                                                                    cohorte_id=cohorte2023.id, status=True,
    #                                                                    requisito__claserequisito__clasificacion=3)[0]
    #                     listf.requisitos = requifinan
    #                     listf.save()
    #
    #     if Rubro.objects.filter(inscripcion=aspirante, status=True).exists():
    #         rubros = Rubro.objects.filter(inscripcion=aspirante, status=True)
    #         for rubro in rubros:
    #             if rubro.admisionposgradotipo == 2:
    #                 qs_anterior = list(Rubro.objects.filter(pk=rubro.id).values())
    #                 tiporubroarancel = TipoOtroRubro.objects.get(pk=2845)
    #                 rubro.nombre = tiporubroarancel.nombre + ' - ' + cohorte2023.maestriaadmision.descripcion + ' - ' + cohorte2023.descripcion
    #                 rubro.cohortemaestria = cohorte2023
    #                 rubro.save()
    #                 # GUARDA AUDITORIA RUBRO
    #                 # qs_nuevo = list(Rubro.objects.filter(pk=rubro.id).values())
    #                 # GUARDA AUDITORIA RUBRO
    #                 if rubro.epunemi and rubro.idrubroepunemi > 0:
    #                     cursor = connections['epunemi'].cursor()
    #                     sql = """UPDATE sagest_rubro SET nombre='%s' WHERE sagest_rubro.status=true and sagest_rubro.id='%s'; """ % (rubro.nombre, rubro.idrubroepunemi)
    #                     # sql = "UPDATE sagest_rubro SET status=false WHERE sagest_rubro.status=true and sagest_rubro.id=" + str(rubro.idrubroepunemi)
    #                     cursor.execute(sql)
    #                     cursor.close()
    #
    #             elif rubro.admisionposgradotipo == 3:
    #                 qs_anterior = list(Rubro.objects.filter(pk=rubro.id).values())
    #                 rubro.nombre = cohorte2023.maestriaadmision.descripcion + ' - ' + cohorte2023.descripcion
    #                 rubro.cohortemaestria = cohorte2023
    #                 rubro.save()
    #                 # GUARDA AUDITORIA RUBRO
    #                 # qs_nuevo = list(Rubro.objects.filter(pk=rubro.id).values())
    #                 # GUARDA AUDITORIA RUBRO
    #                 if rubro.epunemi and rubro.idrubroepunemi > 0:
    #                     cursor = connections['epunemi'].cursor()
    #                     sql = """UPDATE sagest_rubro SET nombre='%s' WHERE sagest_rubro.status=true and sagest_rubro.id='%s'; """ % (rubro.nombre, rubro.idrubroepunemi)
    #                     cursor.execute(sql)
    #                     cursor.close()
    #
    #     aspirante.cohortes_id = cohorte2023.id
    #     aspirante.save()
    #
    #     cont += 1
    #     print(cont, 'Lead:', ' ', aspirante.inscripcionaspirante.persona, 'Cedula:', ' ', aspirante.inscripcionaspirante.persona.cedula,'Cohorte:', ' ', aspirante.cohortes)

    # print(u"********************Neuropsicología***************************")
    #
    # aspirantesmatri = []
    # aspirantes = InscripcionCohorte.objects.filter(status=True, cohortes__id=86)
    #
    # for aspi in aspirantes:
    #     if aspi.inscripcion:
    #         if Matricula.objects.filter(status=True, inscripcion__id=aspi.inscripcion.id).exists():
    #             aspirantesmatri.append(aspi.id)
    #
    # aspirantes = InscripcionCohorte.objects.filter(status=True, cohortes__id=86).exclude(id__in=aspirantesmatri)
    #
    # cont = 0
    # cohorte2023 = CohorteMaestria.objects.get(id=140, status=True)
    # for aspirante in aspirantes:
    #     listarequisitos = EvidenciaRequisitosAspirante.objects.filter(inscripcioncohorte=aspirante)
    #     listarequisitosfinan = EvidenciaRequisitosAspirante.objects.filter(inscripcioncohorte=aspirante,
    #                                                                        requisitos__requisito__claserequisito__clasificacion=3)
    #     for lis in listarequisitos:
    #         if RequisitosMaestria.objects.filter(requisito=lis.requisitos.requisito,
    #                                              cohorte_id=cohorte2023.id, status=True):
    #             requi = RequisitosMaestria.objects.filter(requisito=lis.requisitos.requisito,
    #                                                       cohorte_id=cohorte2023.id,
    #                                                       status=True)[0]
    #             lis.requisitos = requi
    #             lis.save()
    #
    #     if aspirante.formapagopac:
    #         if aspirante.formapagopac.id == 2:
    #             for listf in listarequisitosfinan:
    #                 if RequisitosMaestria.objects.filter(requisito=listf.requisitos.requisito,
    #                                                      cohorte_id=cohorte2023.id, status=True,
    #                                                      requisito__claserequisito__clasificacion=3):
    #                     requifinan = RequisitosMaestria.objects.filter(requisito=listf.requisitos.requisito,
    #                                                                    cohorte_id=cohorte2023.id, status=True,
    #                                                                    requisito__claserequisito__clasificacion=3)[0]
    #                     listf.requisitos = requifinan
    #                     listf.save()
    #
    #     if Rubro.objects.filter(inscripcion=aspirante, status=True).exists():
    #         rubros = Rubro.objects.filter(inscripcion=aspirante, status=True)
    #         for rubro in rubros:
    #             if rubro.admisionposgradotipo == 2:
    #                 qs_anterior = list(Rubro.objects.filter(pk=rubro.id).values())
    #                 tiporubroarancel = TipoOtroRubro.objects.get(pk=2845)
    #                 rubro.nombre = tiporubroarancel.nombre + ' - ' + cohorte2023.maestriaadmision.descripcion + ' - ' + cohorte2023.descripcion
    #                 rubro.cohortemaestria = cohorte2023
    #                 rubro.save()
    #                 # GUARDA AUDITORIA RUBRO
    #                 # qs_nuevo = list(Rubro.objects.filter(pk=rubro.id).values())
    #                 # GUARDA AUDITORIA RUBRO
    #                 if rubro.epunemi and rubro.idrubroepunemi > 0:
    #                     cursor = connections['epunemi'].cursor()
    #                     sql = """UPDATE sagest_rubro SET nombre='%s' WHERE sagest_rubro.status=true and sagest_rubro.id='%s'; """ % (rubro.nombre, rubro.idrubroepunemi)
    #                     # sql = "UPDATE sagest_rubro SET status=false WHERE sagest_rubro.status=true and sagest_rubro.id=" + str(rubro.idrubroepunemi)
    #                     cursor.execute(sql)
    #                     cursor.close()
    #
    #             elif rubro.admisionposgradotipo == 3:
    #                 qs_anterior = list(Rubro.objects.filter(pk=rubro.id).values())
    #                 rubro.nombre = cohorte2023.maestriaadmision.descripcion + ' - ' + cohorte2023.descripcion
    #                 rubro.cohortemaestria = cohorte2023
    #                 rubro.save()
    #                 # GUARDA AUDITORIA RUBRO
    #                 # qs_nuevo = list(Rubro.objects.filter(pk=rubro.id).values())
    #                 # GUARDA AUDITORIA RUBRO
    #                 if rubro.epunemi and rubro.idrubroepunemi > 0:
    #                     cursor = connections['epunemi'].cursor()
    #                     sql = """UPDATE sagest_rubro SET nombre='%s' WHERE sagest_rubro.status=true and sagest_rubro.id='%s'; """ % (rubro.nombre, rubro.idrubroepunemi)
    #                     cursor.execute(sql)
    #                     cursor.close()
    #
    #     aspirante.cohortes_id = cohorte2023.id
    #     aspirante.save()
    #
    #     cont += 1
    #     print(cont, 'Lead:', ' ', aspirante.inscripcionaspirante.persona, 'Cedula:', ' ', aspirante.inscripcionaspirante.persona.cedula,'Cohorte:', ' ', aspirante.cohortes)



# try:
#     aspirantesmatri = []
#     aspirantes = InscripcionCohorte.objects.filter(status=True, cohortes__id=127)
#
#     for aspi in aspirantes:
#         if aspi.inscripcion:
#             if Matricula.objects.filter(status=True, inscripcion__id=aspi.inscripcion.id).exists():
#                 aspirantesmatri.append(aspi.id)
#         elif Rubro.objects.filter(status=True, inscripcion=aspi).exists():
#             firstrubro = Rubro.objects.filter(status=True, inscripcion=aspi).order_by('id')[0]
#             if Pago.objects.filter(status=True, rubro=firstrubro).exists():
#                 aspirantesmatri.append(aspi.id)
#         elif Contrato.objects.filter(status=True, inscripcion=aspi).exists():
#             aspirantesmatri.append(aspi.id)
#
#     aspirantes = InscripcionCohorte.objects.filter(status=True, cohortes__id=127).exclude(id__in=aspirantesmatri)
#
#     cont = 0
#     cohorte2023 = CohorteMaestria.objects.get(id=137, status=True)
#     for aspirante in aspirantes:
#         listarequisitos = EvidenciaRequisitosAspirante.objects.filter(inscripcioncohorte=aspirante)
#         listarequisitosfinan = EvidenciaRequisitosAspirante.objects.filter(inscripcioncohorte=aspirante,
#                                                                            requisitos__requisito__claserequisito__clasificacion=3)
#         for lis in listarequisitos:
#             if RequisitosMaestria.objects.filter(requisito=lis.requisitos.requisito,
#                                                  cohorte_id=cohorte2023.id, status=True):
#                 requi = RequisitosMaestria.objects.filter(requisito=lis.requisitos.requisito,
#                                                           cohorte_id=cohorte2023.id,
#                                                           status=True)[0]
#                 lis.requisitos = requi
#                 lis.save()
#
#         if aspirante.formapagopac:
#             if aspirante.formapagopac.id == 2:
#                 for listf in listarequisitosfinan:
#                     if RequisitosMaestria.objects.filter(requisito=listf.requisitos.requisito,
#                                                          cohorte_id=cohorte2023.id, status=True,
#                                                          requisito__claserequisito__clasificacion=3):
#                         requifinan = RequisitosMaestria.objects.filter(requisito=listf.requisitos.requisito,
#                                                                        cohorte_id=cohorte2023.id, status=True,
#                                                                        requisito__claserequisito__clasificacion=3)[0]
#                         listf.requisitos = requifinan
#                         listf.save()
#
#         if Rubro.objects.filter(inscripcion=aspirante, status=True).exists():
#             rubros = Rubro.objects.filter(inscripcion=aspirante, status=True)
#             for rubro in rubros:
#                 if rubro.admisionposgradotipo == 2:
#                     qs_anterior = list(Rubro.objects.filter(pk=rubro.id).values())
#                     tiporubroarancel = TipoOtroRubro.objects.get(pk=2845)
#                     rubro.nombre = tiporubroarancel.nombre + ' - ' + cohorte2023.maestriaadmision.descripcion + ' - ' + cohorte2023.descripcion
#                     rubro.cohortemaestria = cohorte2023
#                     rubro.save()
#                     # GUARDA AUDITORIA RUBRO
#                     # qs_nuevo = list(Rubro.objects.filter(pk=rubro.id).values())
#                     # GUARDA AUDITORIA RUBRO
#                     if rubro.epunemi and rubro.idrubroepunemi > 0:
#                         cursor = connections['epunemi'].cursor()
#                         sql = """UPDATE sagest_rubro SET nombre='%s' WHERE sagest_rubro.status=true and sagest_rubro.id='%s'; """ % (rubro.nombre, rubro.idrubroepunemi)
#                         # sql = "UPDATE sagest_rubro SET status=false WHERE sagest_rubro.status=true and sagest_rubro.id=" + str(rubro.idrubroepunemi)
#                         cursor.execute(sql)
#                         cursor.close()
#
#                 elif rubro.admisionposgradotipo == 3:
#                     qs_anterior = list(Rubro.objects.filter(pk=rubro.id).values())
#                     rubro.nombre = cohorte2023.maestriaadmision.descripcion + ' - ' + cohorte2023.descripcion
#                     rubro.cohortemaestria = cohorte2023
#                     rubro.save()
#                     # GUARDA AUDITORIA RUBRO
#                     # qs_nuevo = list(Rubro.objects.filter(pk=rubro.id).values())
#                     # GUARDA AUDITORIA RUBRO
#                     if rubro.epunemi and rubro.idrubroepunemi > 0:
#                         cursor = connections['epunemi'].cursor()
#                         sql = """UPDATE sagest_rubro SET nombre='%s' WHERE sagest_rubro.status=true and sagest_rubro.id='%s'; """ % (rubro.nombre, rubro.idrubroepunemi)
#                         cursor.execute(sql)
#                         cursor.close()
#
#         aspirante.cohortes_id = cohorte2023.id
#         aspirante.save()
#
#         cont += 1
#         print(cont, 'Lead:', ' ', aspirante.inscripcionaspirante.persona, 'Cedula:', ' ', aspirante.inscripcionaspirante.persona.cedula,'Cohorte:', ' ', aspirante.cohortes)
#
# except Exception as ex:
#         print('error: %s' % ex)
#
# print("Finaliza")


# try:
#     # leadsbasica = InscripcionCohorte.objects.filter(status=True, cohortes__id=122, inscripcion__isnull=False).values_list('inscripcion__id', flat=True)
#     # matriculados = Matricula.objects.filter(status=True, inscripcion__id__in=leadsbasica, nivel__periodo__id=162)
#
#     matriculados = Matricula.objects.filter(status=True, nivel__periodo__id__in=[182, 160, 167]).order_by('inscripcion__persona__apellido1',
#                                                                                             'inscripcion__persona__apellido2',
#                                                                                             'inscripcion__persona__nombres')
#
#     for matriculado in matriculados:
#         inscrito = InscripcionCohorte.objects.filter(status=True, inscripcion=matriculado.inscripcion).first()
#         if inscrito is not None:
#             rubrosinscrito = Rubro.objects.filter(status=True, inscripcion=inscrito,persona=inscrito.inscripcionaspirante.persona)
#             for rubro in rubrosinscrito:
#                 if not rubro.matricula:
#                     rubro.matricula = matriculado
#                     rubro.save()
#                     print(f"Se ha actualizado rubro de {inscrito.inscripcionaspirante.persona} {rubro.id} - {rubro.matricula}")
#                 print(f"Ya tiene matricula {inscrito.inscripcionaspirante.persona}  {rubro.id} - {rubro.matricula}")
#         else:
#             print(f"NO ID INSCRITO {matriculado.inscripcion.persona} - {matriculado.inscripcion.persona.cedula}")
# except Exception as ex:
#         print('error: %s' % ex)
#
# print("Finaliza")




# try:
#     miarchivo = openpyxl.load_workbook("contabilidad_financiamiento.xlsx")
#     lista = miarchivo.get_sheet_by_name('resultados')
#     totallista = lista.rows
#     a=0
#     c=0
#     for filas in totallista:
#         a += 1
#         if a > 1:
#             idinscripcioncohorte = str(filas[0].value)
#             # cedula = str((filas[6].value).replace(".", ""))
#             cedula = str((filas[6].value)).replace(".", "")
#             tipo = TipoFormaPagoPac.objects.get(id=2, status=True)
#             if idinscripcioncohorte != 'None':
#                 observacion = '30%+10 CUOTAS'
#                 if InscripcionCohorte.objects.filter(status=True, id=idinscripcioncohorte).exists():
#                     inscripcion = InscripcionCohorte.objects.get(status=True, id=idinscripcioncohorte)
#                     if AsesorComercial.objects.filter(status=True, persona__cedula=cedula).exists():
#                         asesor = AsesorComercial.objects.get(status=True, persona__cedula=cedula)
#
#                         inscripcion.formapagopac = tipo
#                         inscripcion.estadoformapago = 1
#                         inscripcion.save()
#
#                         deta = DetalleAprobacionFormaPago(inscripcion_id=inscripcion.id,
#                                                           formapagopac=tipo,
#                                                           estadoformapago=1,
#                                                           observacion=observacion,
#                                                           persona=asesor.persona)
#                         deta.save()
#
#                         c += 1
#                         print(c,' ','Lead:',inscripcion.inscripcionaspirante.persona,' - ','Asesor:',asesor, ' - ', 'FormaPago:', inscripcion.formapagopac.descripcion)
#                         # else:
#                         #     print(u"YA TIENE ASESOR:",'-', inscripcion.inscripcionaspirante.persona)
#                     else:
#                         print(u" NO EXISTE EL ASESOR", inscripcion.inscripcionaspirante.persona)
#                 else:
#                     print(u"-NO EXISTE EL LEAD")
#             else:
#                 print(u"FINALIZADO")
#                 break
# except Exception as ex:
#         print('error: %s' % ex)
#
# print("Finaliza")

# try:
#     # cohorte = CohorteMaestria.objects.get(id__in=[116, 118]).values_list('id')
#     id_inscripcion = Matricula.objects.filter(inscripcion__carrera__coordinacion__id=7).values_list('inscripcion__id', flat=True)
#
#     prospectos = InscripcionCohorte.objects.filter(status=True, cohortes__maestriaadmision__carrera__coordinacion__id=7,
#                            fecha_creacion__in=InscripcionCohorte.objects.values('inscripcionaspirante__id').annotate(fecha_creacion=Max('fecha_creacion')).values_list('fecha_creacion', flat=True).filter(status=True),
#                                                    estado_aprobador=2, fecha_creacion__lte = '2022-08-31').exclude(inscripcion__id__in=id_inscripcion)
#     tipo = TipoRespuestaProspecto.objects.get(id=5, status=True)
#     cont = 0
#     for prospecto in prospectos:
#         aspi = InscripcionCohorte.objects.get(pk=prospecto.id, status=True)
#
#         aspi.tiporespuesta = tipo
#         aspi.save()
#         cont += 1
#
#         print(cont, 'Lead:', ' ', aspi.inscripcionaspirante.persona, 'Respuesta:', ' ', aspi.tiporespuesta.descripcion)
#
# except Exception as ex:
#         print('error: %s' % ex)
#
# print("Finaliza")


# try:
#     # cohorte = CohorteMaestria.objects.get(id__in=[116, 118]).values_list('id')
#     id_inscripcion = Matricula.objects.filter(inscripcion__carrera__coordinacion__id=7).values_list('inscripcion__id', flat=True)
#
#     prospectos = InscripcionCohorte.objects.filter(status=True, cohortes__maestriaadmision__carrera__coordinacion__id=7,
#                            fecha_creacion__in=InscripcionCohorte.objects.values('inscripcionaspirante__id').annotate(fecha_creacion=Max('fecha_creacion')).values_list('fecha_creacion', flat=True).filter(status=True),
#                                                    estado_aprobador__in=[1, 3], fecha_creacion__lte = '2022-08-31').exclude(inscripcion__id__in=id_inscripcion, cohortes__id__in=[120, 123, 122, 113, 121, 86, 124, 107, 125, 126, 127, 128])
#     tipo = TipoRespuestaProspecto.objects.get(id=4, status=True)
#     cont = 0
#     for prospecto in prospectos:
#         aspi = InscripcionCohorte.objects.get(pk=prospecto.id, status=True)
#
#         aspi.tiporespuesta = tipo
#         aspi.save()
#         cont += 1
#
#         print(cont, 'Lead:', ' ', aspi.inscripcionaspirante.persona, 'Respuesta:', ' ', aspi.tiporespuesta.descripcion)
#
# except Exception as ex:
#         print('error: %s' % ex)
#
# print("Finaliza")

# try:
#     aspirantesmatri = []
#     aspirantes = InscripcionCohorte.objects.filter(status=True, id__in=[28791, 21682, 27463, 19158, 27692, 28231, 25270, 28701, 26615, 25141])
#
#     # for aspi in aspirantes:
#     #     if aspi.inscripcion:
#     #         if Matricula.objects.filter(status=True, inscripcion__id=aspi.inscripcion.id).exists():
#     #             aspirantesmatri.append(aspi.id)
#
#     # aspirantes = InscripcionCohorte.objects.filter(status=True, cohortes__id=122).exclude(id__in=aspirantesmatri)
#
#     cont = 0
#     cohorte2023 = CohorteMaestria.objects.get(id=122, status=True)
#     for aspirante in aspirantes:
#         listarequisitos = EvidenciaRequisitosAspirante.objects.filter(inscripcioncohorte=aspirante)
#         listarequisitosfinan = EvidenciaRequisitosAspirante.objects.filter(inscripcioncohorte=aspirante,
#                                                                            requisitos__requisito__claserequisito__clasificacion=3)
#         for lis in listarequisitos:
#             if RequisitosMaestria.objects.filter(requisito=lis.requisitos.requisito,
#                                                  cohorte_id=cohorte2023.id, status=True):
#                 requi = RequisitosMaestria.objects.filter(requisito=lis.requisitos.requisito,
#                                                           cohorte_id=cohorte2023.id,
#                                                           status=True)[0]
#                 lis.requisitos = requi
#                 lis.save()
#
#         if aspirante.formapagopac:
#             if aspirante.formapagopac.id == 2:
#                 for listf in listarequisitosfinan:
#                     if RequisitosMaestria.objects.filter(requisito=listf.requisitos.requisito,
#                                                          cohorte_id=cohorte2023.id, status=True,
#                                                          requisito__claserequisito__clasificacion=3):
#                         requifinan = RequisitosMaestria.objects.filter(requisito=listf.requisitos.requisito,
#                                                                        cohorte_id=cohorte2023.id, status=True,
#                                                                        requisito__claserequisito__clasificacion=3)[0]
#                         listf.requisitos = requifinan
#                         listf.save()
#
#         if Rubro.objects.filter(inscripcion=aspirante, status=True).exists():
#             rubros = Rubro.objects.filter(inscripcion=aspirante, status=True)
#             for rubro in rubros:
#                 if rubro.admisionposgradotipo == 2:
#                     qs_anterior = list(Rubro.objects.filter(pk=rubro.id).values())
#                     tiporubroarancel = TipoOtroRubro.objects.get(pk=2845)
#                     rubro.nombre = tiporubroarancel.nombre + ' - ' + cohorte2023.maestriaadmision.descripcion + ' - ' + cohorte2023.descripcion
#                     rubro.cohortemaestria = cohorte2023
#                     rubro.save()
#                     # GUARDA AUDITORIA RUBRO
#                     # qs_nuevo = list(Rubro.objects.filter(pk=rubro.id).values())
#                     # GUARDA AUDITORIA RUBRO
#                     if rubro.epunemi and rubro.idrubroepunemi > 0:
#                         cursor = connections['epunemi'].cursor()
#                         sql = "UPDATE sagest_rubro SET status=false WHERE sagest_rubro.status=true and sagest_rubro.id=" + str(rubro.idrubroepunemi)
#                         cursor.execute(sql)
#                         cursor.close()
#
#                 elif rubro.admisionposgradotipo == 3:
#                     qs_anterior = list(Rubro.objects.filter(pk=rubro.id).values())
#                     rubro.nombre = cohorte2023.maestriaadmision.descripcion + ' - ' + cohorte2023.descripcion
#                     rubro.cohortemaestria = cohorte2023
#                     rubro.save()
#                     # GUARDA AUDITORIA RUBRO
#                     # qs_nuevo = list(Rubro.objects.filter(pk=rubro.id).values())
#                     # GUARDA AUDITORIA RUBRO
#                     if rubro.epunemi and rubro.idrubroepunemi > 0:
#                         cursor = connections['epunemi'].cursor()
#                         sql = "UPDATE sagest_rubro SET status=false WHERE sagest_rubro.status=true and sagest_rubro.id=" + str(rubro.idrubroepunemi)
#                         cursor.execute(sql)
#                         cursor.close()
#
#         aspirante.cohortes_id = cohorte2023.id
#         aspirante.save()
#
#         cont += 1
#         print(cont, 'Lead:', ' ', aspirante.inscripcionaspirante.persona, 'Cedula:', ' ', aspirante.inscripcionaspirante.persona.cedula,'Cohorte:', ' ', aspirante.cohortes)
#
# except Exception as ex:
#         print('error: %s' % ex)
#
# print("Finaliza")


# try:
#     miarchivo = openpyxl.load_workbook("masivo_finanaciamiento_original.xlsx")
#     lista = miarchivo.get_sheet_by_name('resultados')
#     totallista = lista.rows
#     a=0
#     c=0
#     for filas in totallista:
#         a += 1
#         if a > 1:
#             idinscripcioncohorte = str(filas[1].value)
#             # cedula = str((filas[6].value).replace(".", ""))
#             cedula = str((filas[7].value)).replace(".", "")
#             tipo = TipoFormaPagoPac.objects.get(id=2, status=True)
#             if idinscripcioncohorte != 'None':
#                 if str(filas[9].value) == 'None':
#                     observacion = 'NINGUNA'
#                 else:
#                     observacion = str(filas[9].value)
#                 if InscripcionCohorte.objects.filter(status=True, id=idinscripcioncohorte).exists():
#                     inscripcion = InscripcionCohorte.objects.get(status=True, id=idinscripcioncohorte)
#                     if AsesorComercial.objects.filter(status=True, persona__cedula=cedula).exists():
#                         asesor = AsesorComercial.objects.get(status=True, persona__cedula=cedula)
#
#                         inscripcion.formapagopac = tipo
#                         inscripcion.estadoformapago = 1
#                         inscripcion.save()
#
#                         deta = DetalleAprobacionFormaPago(inscripcion_id=inscripcion.id,
#                                                           formapagopac=tipo,
#                                                           estadoformapago=1,
#                                                           observacion=observacion,
#                                                           persona=asesor.persona)
#                         deta.save()
#
#                         c += 1
#                         print(c,' ','Lead:',inscripcion.inscripcionaspirante.persona,' - ','Asesor:',asesor, ' - ', 'FormaPago:', inscripcion.formapagopac.descripcion, ' - ', observacion)
#                         # else:
#                         #     print(u"YA TIENE ASESOR:",'-', inscripcion.inscripcionaspirante.persona)
#                     else:
#                         print(u" NO EXISTE EL ASESOR", inscripcion.inscripcionaspirante.persona)
#                 else:
#                     print(u"-NO EXISTE EL LEAD")
#             else:
#                 print(u"FINALIZADO")
#                 break
# except Exception as ex:
#         print('error: %s' % ex)
#
# print("Finaliza")


# try:
#     miarchivo = openpyxl.load_workbook("contabilidad_financiamiento.xlsx")
#     lista = miarchivo.get_sheet_by_name('ventas')
#     totallista = lista.rows
#     a=0
#     for filas in totallista:
#         a += 1
#         if a > 1:
#             cedula = str(filas[9].value)
#             if cedula == 'None':
#                 cedula = ''
#             else:
#                 cedula = str((filas[9].value).replace(".", ""))
#             idinscripcioncohorte = str(filas[1].value)
#             if idinscripcioncohorte != 'None':
#                 observacion = 'MASIVO VENTAS COHORTE II BASICA'
#                 # urlzoom=u"https://unemi-edu-ec.zoom.us/j/%s"%idzoom
#                 # print(u"%s"%correo)
#                 if InscripcionCohorte.objects.filter(status=True, id=idinscripcioncohorte).exists():
#                     inscripcion = InscripcionCohorte.objects.get(status=True, id=idinscripcioncohorte)
#                     asesoranti = inscripcion.asesor
#                     if AsesorComercial.objects.filter(status=True,persona__cedula=cedula).exists():
#                         asesor = AsesorComercial.objects.get(status=True, persona__cedula=cedula)
#                         # if not inscripcion.asesor:
#                         inscripcion.asesor = asesor
#                         inscripcion.estado_asesor = 2
#                         inscripcion.save()
#
#                         if not asesoranti:
#                             histo = HistorialAsesor(inscripcion_id=inscripcion.id, fecha_inicio=inscripcion.fecha_modificacion,
#                                             fecha_fin=None, asesor=inscripcion.asesor, observacion=observacion)
#                             histo.save()
#                         else:
#                             if not asesoranti == asesor:
#                                 histoanti = HistorialAsesor.objects.get(inscripcion_id=inscripcion.id, fecha_fin=None)
#                                 histoanti.fecha_fin = inscripcion.fecha_modificacion
#                                 histoanti.save()
#                                 histo = HistorialAsesor(inscripcion_id=inscripcion.id, fecha_inicio=inscripcion.fecha_modificacion,
#                                                         fecha_fin=None, asesor=inscripcion.asesor, observacion=observacion)
#                                 histo.save()
#                             else:
#                                 print('El asesor ya tiene este lead asignado')
#
#                         print('Lead:',inscripcion.inscripcionaspirante.persona,' - ','Asesor:',asesor)
#                         # else:
#                         #     print(u"YA TIENE ASESOR:",'-', inscripcion.inscripcionaspirante.persona)
#                     else:
#                         print(u" NO EXISTE EL ASESOR", inscripcion.inscripcionaspirante.persona)
#                 else:
#                     print(u"-NO EXISTE EL LEAD")
#             else:
#                 print(u"FINALIZADO")
#                 break
# except Exception as ex:
#         print('error: %s' % ex)
#
# print("Finaliza")