# -*- coding: UTF-8 -*-
import json
from datetime import datetime, timedelta, date
from decimal import Decimal
import os
import io
import collections
import calendar
import sys

from django.core.exceptions import ObjectDoesNotExist
from django.db.models import Avg, Q
from django.contrib import messages
from django.forms import model_to_dict
from django.contrib.contenttypes.models import ContentType
from unidecode import unidecode

from core.firmar_documentos import firmararchivogenerado, obtener_posicion_x_y_saltolinea
from core.firmar_documentos_ec import JavaFirmaEc
from core.choices.models.sagest import ESTADO_INCIDENCIA, ETAPA_INCIDENCIA

from django.core.files.base import ContentFile

from core.firmar_documentos_ec_descentralizada import qrImgFirma
from sagest.commonviews import obtener_estado_solicitud
from sagest.th_marcadas import calculando_marcadas
from sga.funcionesxhtml2pdf import conviert_html_to_pdfsave, conviert_html_to_pdf_save_informe, \
    conviert_html_to_pdfsaveqr_generico, conviert_html_to_pdf_name_bitacora, conviert_html_to_pdfsave_generic_lotes
import pyqrcode
import fitz.utils
import code128
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.http import HttpResponseRedirect, JsonResponse, HttpResponse
from django.shortcuts import render
from django.template.loader import get_template
from decorators import secure_module, last_access
from med.models import PersonaExamenFisico
from sga.funcionesxhtml2pdf import conviert_html_to_pdf, conviert_html_to_pdf_name
from sagest.forms import DatosPersonalesForm, DatosNacimientoForm, DatosMedicosForm, DatosDomicilioForm, \
    ContactoEmergenciaForm, EtniaForm, DiscapacidadForm, FamiliarForm, DeclaracionBienForm, CuentaBancariaPersonaForm, \
    TitulacionPersonaForm, ArchivoTitulacionForm, CapacitacionPersonaForm, ArchivoCapacitacionForm, \
    ExperienciaLaboralForm, ArchivoIdiomaForm, ArchivoExperienciaForm, DocumentoInscripcionForm, \
    DatosInstitucionalesPersonaForm, DetalleTitulacionBachillerForm, OtroMeritoForm, CertificadoPersonaForm, \
    InformeMensualForm, ArticuloInvestigacionForm, DatosBecaForm, InstitucionForm, DeportistaForm, MigranteForm, \
    ArtistaForm, BitacoraForm, VacunacionCovidForm, VacunacionCovidEvidenciaForm, TituloHojaVidaForm, \
    CambiarMarcadaForm, DatosSituacionLaboralForm, PersonaTransporteForm, \
    PersonaEnfermedadForm, ArchivoHistorialContratoForm, AccionPersonalDocumentoForm, SolicitudJustificacionMarcadaForm, \
    PersonaAlimentacionUniversidadForm, PersonaPlanTelefonicoForm, PersonaCompraAlimentosForm, PersonaGastoMensualForm, \
    PersonaDetalleMaternidadAspiranteForm, DatosPersonalesAspiranteForm, DatosNacimientoAspiranteForm, \
    DatosDomicilioAdmisionForm, \
    DatosSituacionLaboralAspiranteForm, ContactoEmergenciaAspiranteForm, EtniaAspiranteForm, DiscapacidadAspiranteForm, \
    FamiliarAspiranteForm, DeclaracionBienAspiranteForm, VacunacionCovidAdmisionForm, \
    VacunacionCovidEvidenciaAdmisionForm, \
    TitulacionPersonaAdmisionForm, TituloHojaVidaAdmisionForm, PazSalvoFormHV, PersonaDetalleMaternidadForm, \
    RedAcademicaForm, ParRevisorArticuloForm, ProyectoInvestigacionExternoForm, RegistraDecimoForm, SubirRequisitoPSForm

from sagest.models import ExperienciaLaboral, RolPago, ActivoFijo, GastosPersonales, PeriodoGastosPersonales, \
    SolicitudPublicacion, ResumenMesGastosPersonales, CapEventoPeriodo, CapCabeceraSolicitud, CapDetalleSolicitud, \
    Departamento, DistributivoPersona, TrabajadorDiaJornada, OtroMerito, LogDia, ParticipanteSolicitudPublicacion, \
    Formulario107, BitacoraActividadDiaria, VacunaCovid, VacunaCovidDosis, LogMarcada, MarcadasDia, RegistroMarcada, \
    SolicitudActivos, HistorialArchivosContratos, PersonaContratos, PersonaAcciones, \
    HistoricoDocumentosPersonaAcciones, SolicitudJustificacionMarcada, DetalleSolicitudJustificacionMarcada, \
    HistorialSolicitudJustificacionMarcada, AccionPersonal, TIPO_SISTEMA, TIPO_ACTIVIDAD_BITACORA, PazSalvo, \
    HistorialCertificadoFirmaPS, FormatoPazSalvo, DenominacionPuesto, \
    DistributivoPersonaHistorial, RegistroDecimo, ConfiguraDecimo, RequisitoPazSalvo, DocumentoPazSalvo, RequisitoPeriodotthh, ConfiguraPeriodotthh, \
    PersonaPeriodotthh, DocumentoPersonaPeriodotthh
from settings import SITE_STORAGE, PUESTO_ACTIVO_ID, MEDIA_ROOT, DEBUG
from settings import PORCENTAJE_SEGURO
from sga.commonviews import adduserdata, obtener_reporte
from sga.forms import IdiomaDominaForm, SolicitudPublicacionForm, EvidenciaForm, ReferenciaPersonaForm, \
    PersonaSubirFotoForm, RevistaInvestigacionForm, \
    RevistaInvestigacionHojaVidaForm, ParticipanteProfesorArticuloForm, ParticipanteAdministrativoArticuloForm, \
    ParticipanteInscripcionArticuloForm, SolicitudPublicacionLibroForm, \
    CertificadoIdiomaForm, CapituloLibroHojaVidaForm, CertificadoTutoriaForm, \
    PersonaAportacionHistorialLaboralForm, PersonaAportacionHistorialLaboralAspiranteForm
from sga.funciones import generar_nombre, log, MiPaginador, tituloinstitucion, variable_valor, convertir_fecha, \
    convertir_fecha_hora, convertir_fecha_invertida, email_valido, null_to_decimal, puede_realizar_accion_is_superuser, \
    validar_archivo, puede_realizar_accion, notificacion as noti, remover_caracteres_especiales_unicode, \
    convertir_fecha_hora_invertida, \
    elimina_tildes, puede_realizar_accion_afirmativo, validarcedula
from sga.models import PersonaDatosFamiliares, DeclaracionBienes, CuentaBancariaPersona, Titulacion, Capacitacion, \
    TemaTitulacionPosgradoMatricula, \
    IdiomaDomina, TIPO_CELULAR, ParticipantesMatrices, ArticuloInvestigacion, PonenciasInvestigacion, \
    LibroInvestigacion, null_to_numeric, Evidencia, NivelTitulacion, Persona, RespuestaEvaluacionAcreditacion, \
    ResumenFinalProcesoEvaluacionIntegral, MigracionEvaluacionDocente, ResumenFinalEvaluacionAcreditacion, \
    ResumenParcialEvaluacionIntegral, CapituloLibroInvestigacion, ResponsableEvaluacion, Inscripcion, Archivo, \
    DetalleTitulacionBachiller, CUENTAS_CORREOS, ReferenciaPersona, CertificacionPersona, InformeMensual, MESES_CHOICES, \
    Coordinacion, Materia, AsignaturaMallaPreferencia, AsignaturaMalla, Carrera, Malla, NivelMalla, Silabo, Profesor, \
    ProyectoInvestigacionExterno, BecaPersona, InstitucionBeca, DeportistaPersona, MigrantePersona, ArtistaPersona, \
    BecaAsignacion, PersonaDocumentoPersonal, FotoPersona, RevistaInvestigacion, RevistaInvestigacionBase, \
    BaseIndexadaInvestigacion, TIPO_PARTICIPANTE_INSTITUCION, ProcesoEvaluativoAcreditacion, RubricaPreguntas, \
    CapEventoPeriodoDocente, CapCabeceraSolicitudDocente, CapDetalleSolicitudDocente, Titulo, \
    ParRevisorProduccionCientifica, PeriodoPromocionDocente, RedPersona, CertificadoIdioma, CertificadoTutoriaHV, \
    PersonaAportacionHistorialLaboral, SubAreaConocimientoTitulacion, SubAreaEspecificaConocimientoTitulacion, \
    CamposTitulosPostulacion, AreaConocimientoTitulacion, Graduado, SubTipoDiscapacidad, PersonaSituacionLaboral, \
    PersonaTransporte, PersonaAlimentacionUniversidad, PersonaPlanTelefonico, PersonaCompraAlimentos, \
    PersonaGastoMensual, PersonaEnfermedad, Notificacion, ProfesorMateria, CoordinadorCarrera, ClaseActividad, \
    Externo, PerfilUsuario, PersonaDetalleMaternidad, EvidenciaActividadDetalleDistributivo
from sga.tasks import send_html_mail
from sga.templatetags.sga_extras import encrypt
from sagest.funciones import dominio_sistema_base, encrypt_id, ext_archive
from xlwt import *
import random
import openpyxl
from pdip.models import ContratoDip, HistorialPagoMes, ActaPago, InformeTecnico, InformeActividadJornada, \
    MemoActividadPosgrado, SecuenciaMemoActividadPosgrado, \
    PlantillaContratoDip, ActividadesPerfil, ContratoCarrera, SolicitudPago, RequisitoSolicitudPago, \
    HistorialProcesoSolicitud, ContratoAreaPrograma, GrupoRevisionPagoContrato, HistorialObseracionSolicitudPago
from pdip.forms import ArchivoInformesForm
from helpdesk.models import SolicitudConfirmacionMantenimiento, HistorialSolicitudConfirmacionMantenimiento
from utils.filtros_genericos import filtro_persona_select
from django.core.files import File as DjangoFile

#APP DIRECTIVO
from directivo.models import PersonaSancion, IncidenciaSancion, RespuestaDescargo, AudienciaSancion, PersonaAudienciaSancion, ConsultaFirmaPersonaSancion
from directivo.utils.funciones import secciones_etapa_analisis, secciones_etapa_audiencia, firmar_documento_etapa
from directivo.forms import RespuestaDescargoForm, JustificacionPersonaAudienciaForm, MotivoNoFirmaAccionPersonalForm
from directivo.utils.actions_sanciones import get_revisar_audiencia
unicode = str


@login_required(redirect_field_name='ret', login_url='/loginsagest')
@secure_module
@last_access
@transaction.atomic()
def view(request):
    data = {}
    adduserdata(request, data)
    persona = request.session['persona']
    periodoacademico = request.session.get('periodo')
    perfilprincipal = request.session['perfilprincipal']
    data['profesor'] = profesor = persona.profesor()
    data['puede_modificar_hv'] = variable_valor('PUEDE_MODIFICAR_HV')
    esestudiante = True if perfilprincipal.es_estudiante() else False
    data['DOMINIO_DEL_SISTEMA'] = dominio_sistema = dominio_sistema_base(request)
    data['pendiente_hijos'] = True if (persona.datos_extension().tienehijos == 1 and
                                       persona.distributivopersona_set.filter().exists()) else False
    hoy = datetime.now()
    if request.method == 'POST':
        action = request.POST['action']

        if action == 'validarLenguaMaterna':
            try:
                idioma = IdiomaDomina.objects.get(pk=request.POST['ididioma'])
                if idioma.lenguamaterna:
                    idioma.lenguamaterna = False
                else:
                    if persona.idiomadomina_set.filter(status=True, lenguamaterna=True).exclude(pk=idioma.pk).exists():
                        return JsonResponse({'result': 'bad', 'mensaje': u'Ya existe una lengua materna registrada.'})
                    idioma.lenguamaterna = True
                idioma.save(request)
                return JsonResponse({'result': 'ok'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'datospersonalesaspirante':
            try:
                persona = request.session['persona']
                persona = Persona.objects.get(id=persona.id)
                f = DatosPersonalesAspiranteForm(request.POST, request.FILES)
                if not f.is_valid():
                    transaction.set_rollback(True)
                    return JsonResponse({'result': 'bad', "form": [{k: v[0]} for k, v in f.errors.items()],
                                         "mensaje": "Error en el formulario"})
                if esestudiante:
                    f.es_estudiante()
                f_maternidad = PersonaDetalleMaternidadAspiranteForm(request.POST)

                if 'archivocedula' in request.FILES:
                    arch = request.FILES['archivocedula']
                    extension = arch._name.split('.')
                    tam = len(extension)
                    exte = extension[tam - 1]
                    if arch.size > 4194304:
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"Error, el tamaño del archivo es mayor a 4 Mb."})
                    if not exte.lower() == 'pdf':
                        return JsonResponse({"result": "bad", "mensaje": u"Solo se permiten archivos .pdf"})
                if f.is_valid():
                    persona.pasaporte = f.cleaned_data['pasaporte']
                    persona.anioresidencia = f.cleaned_data['anioresidencia']
                    persona.nacimiento = f.cleaned_data['nacimiento']
                    if not esestudiante:
                        persona.telefonoextension = f.cleaned_data['extension']
                    # persona.nacionalidad = f.cleaned_data['nacionalidad']
                    persona.sexo = f.cleaned_data['sexo']
                    persona.lgtbi = f.cleaned_data['lgtbi']
                    persona.email = f.cleaned_data['email']
                    persona.libretamilitar = f.cleaned_data['libretamilitar']
                    persona.eszurdo = f.cleaned_data['eszurdo']
                    persona.save(request)
                    personaextension = persona.datos_extension()
                    personaextension.estadocivil = f.cleaned_data['estadocivil']
                    personaextension.save(request)

                    if 'archivocedula' in request.FILES:
                        newfile = request.FILES['archivocedula']
                        newfile._name = generar_nombre("cedula", newfile._name)

                        documento = persona.documentos_personales()
                        if documento is None:
                            documento = PersonaDocumentoPersonal(persona=persona,
                                                                 cedula=newfile,
                                                                 estadocedula=1
                                                                 )
                        else:
                            documento.cedula = newfile
                            documento.estadocedula = 1

                        documento.save(request)

                    if 'papeleta' in request.FILES:
                        newfile = request.FILES['papeleta']
                        newfile._name = generar_nombre("papeleta", newfile._name)

                        documento = persona.documentos_personales()
                        if documento is None:
                            documento = PersonaDocumentoPersonal(persona=persona,
                                                                 papeleta=newfile,
                                                                 estadopapeleta=1
                                                                 )
                        else:
                            documento.papeleta = newfile
                            documento.estadopapeleta = 1

                        documento.save(request)

                    if 'archivolibretamilitar' in request.FILES:
                        newfile = request.FILES['archivolibretamilitar']
                        newfile._name = generar_nombre("libretamilitar", newfile._name)

                        documento = persona.documentos_personales()
                        if documento is None:
                            documento = PersonaDocumentoPersonal(persona=persona,
                                                                 libretamilitar=newfile,
                                                                 estadolibretamilitar=1
                                                                 )
                        else:
                            documento.libretamilitar = newfile
                            documento.estadolibretamilitar = 1

                        documento.save(request)

                if f_maternidad.is_valid():
                    if f_maternidad.cleaned_data['estadogestacion']:
                        personamaternidad = persona.maternidad()
                        personamaternidad.gestacion = f_maternidad.cleaned_data['estadogestacion']
                        personamaternidad.semanasembarazo = f_maternidad.cleaned_data['semanasembarazo']
                        personamaternidad.status_gestacion = True
                        personamaternidad.save(request)
                    if not f_maternidad.cleaned_data['estadogestacion']:
                        if f_maternidad.cleaned_data['lactancia']:
                            personamaternidad = persona.maternidad()
                            personamaternidad.lactancia = f_maternidad.cleaned_data['lactancia']
                            personamaternidad.fechaparto = f_maternidad.cleaned_data['fechaparto']
                            personamaternidad.status_lactancia = True
                            personamaternidad.save(request)
                        if persona.personadetallematernidad_set.filter(Q(status_gestacion=True)).exists():
                            personamaternidad = persona.personadetallematernidad_set.filter(
                                Q(status_gestacion=True)).last()
                            personamaternidad.status_gestacion = False
                            personamaternidad.save(request)
                    if not f_maternidad.cleaned_data['lactancia']:
                        if persona.personadetallematernidad_set.filter(Q(status_lactancia=True)).exists():
                            personamaternidad = persona.personadetallematernidad_set.filter(
                                Q(status_lactancia=True)).last()
                            personamaternidad.status_lactancia = False
                            personamaternidad.save(request)
                    log(u'Modifico datos personales: %s' % persona, request, "edit")
                    return JsonResponse({'result': 'ok'})
                else:
                    raise NameError('Error en el formulario')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', 'mensaje': f'Error al guardar los datos: {ex}'})

        elif action == 'datossituacionlaboralaspirante':
            try:
                persona = request.session['persona']
                if PersonaSituacionLaboral.objects.filter(persona_id=persona.id, status=True).exists():
                    situacionlaboral = PersonaSituacionLaboral.objects.get(persona_id=persona.id)
                else:
                    situacionlaboral = PersonaSituacionLaboral(persona=persona)
                    situacionlaboral.save(request)
                f = DatosSituacionLaboralAspiranteForm(request.POST)
                if f.is_valid():
                    situacionlaboral.tipoinstitucionlaboral = 0
                    situacionlaboral.lugartrabajo = ''
                    situacionlaboral.negocio = ''
                    situacionlaboral.disponetrabajo = f.cleaned_data['disponetrabajo']
                    if situacionlaboral.disponetrabajo:
                        situacionlaboral.tipoinstitucionlaboral = f.cleaned_data['tipoinstitucionlaboral']
                        situacionlaboral.lugartrabajo = f.cleaned_data['lugartrabajo']
                    situacionlaboral.buscaempleo = f.cleaned_data['buscaempleo']
                    situacionlaboral.tienenegocio = f.cleaned_data['tienenegocio']
                    if situacionlaboral.tienenegocio:
                        situacionlaboral.negocio = f.cleaned_data['negocio']
                    situacionlaboral.save(request)

                    log(u'Modifico datos de situacion laboral: %s' % persona, request, "datossituacionlaboral")
                    return JsonResponse({'result': 'ok'})

            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', 'mensaje': u'Error al guardar los datos'})

        elif action == 'addvacunacvd19aspirante':
            try:
                with transaction.atomic():
                    newfile = None
                    vacunado = False
                    form = VacunacionCovidAdmisionForm(request.POST, request.FILES)
                    if form.is_valid():
                        vacunacion = VacunaCovid(persona=persona)
                        if 'recibiovacuna' in request.POST:
                            vacunado = True if int(request.POST['recibiovacuna']) == 1 else False
                            vacunacion.recibiovacuna = vacunado
                        if 'recibiodosiscompleta' in request.POST:
                            recibiodosiscompleta = True if int(request.POST['recibiodosiscompleta']) == 1 else False
                            vacunacion.recibiodosiscompleta = recibiodosiscompleta
                        if 'deseavacunarse' in request.POST:
                            deseavacunarse = True if int(request.POST['deseavacunarse']) == 1 else False
                            vacunacion.deseavacunarse = deseavacunarse
                        if 'tipovacuna' in request.POST:
                            vacunacion.tipovacuna = form.cleaned_data['tipovacuna'] if form.cleaned_data[
                                'tipovacuna'] else None
                        # if 'fecha_certificado' in request.POST:
                        #     vacunacion.fecha_certificado = form.cleaned_data['fecha_certificado'] if form.cleaned_data['fecha_certificado'] else None
                        if 'deseavacunarse' in request.POST:
                            vacunacion.deseavacunarse = True if int(request.POST['deseavacunarse']) == 1 else False
                        if 'certificado' in request.FILES:
                            newfile = request.FILES['certificado']
                            extension = newfile._name.split('.')
                            tam = len(extension)
                            exte = extension[tam - 1]
                            if newfile.size > 4194304:
                                return JsonResponse(
                                    {"result": True, "mensaje": u"Error, el tamaño del archivo es mayor a 4 Mb."})
                            if exte in ['pdf', 'jpg', 'jpeg', 'png', 'jpeg', 'peg']:
                                newfile._name = generar_nombre("certificado_", newfile._name)
                            else:
                                return JsonResponse(
                                    {"result": True, "mensaje": u"Error, solo archivos .pdf,.jpg, .jpeg"})
                            vacunacion.certificado = newfile
                        vacunacion.save(request)
                        if vacunado:
                            datosdosis = request.POST.getlist('infoDosis[]')
                            if datosdosis:
                                c = 0
                                while c < len(datosdosis):
                                    dosis = VacunaCovidDosis(cabvacuna=vacunacion, numdosis=datosdosis[c],
                                                             fechadosis=datosdosis[c + 1])
                                    dosis.save(request)
                                    c += 2
                            else:
                                transaction.set_rollback(True)
                                return JsonResponse({"result": True, "mensaje": "Registre al menos una dosis."},
                                                    safe=False)

                        log(u'Adiciono evidencia de vacunación: %s' % vacunacion, request, "add")
                        return JsonResponse({"result": False, 'to': request.path}, safe=False)
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({"result": True, "mensaje": "Complete los datos requeridos."}, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        elif action == 'addcertificadovacunacovidaspirante':
            try:
                with transaction.atomic():
                    newfile = None
                    vacunado = False
                    form = VacunacionCovidEvidenciaAdmisionForm(request.POST, request.FILES)
                    if form.is_valid():
                        vacunacion = VacunaCovid.objects.get(pk=int(request.POST['id']))
                        # if 'fecha_certificado' in request.POST:
                        #     vacunacion.fecha_certificado = form.cleaned_data['fecha_certificado'] if form.cleaned_data['fecha_certificado'] else None
                        if 'certificado' in request.FILES:
                            newfile = request.FILES['certificado']
                            extension = newfile._name.split('.')
                            tam = len(extension)
                            exte = extension[tam - 1]
                            if newfile.size > 4194304:
                                return JsonResponse(
                                    {"result": True, "mensaje": u"Error, el tamaño del archivo es mayor a 4 Mb."})
                            if exte in ['pdf', 'jpg', 'jpeg', 'png', 'jpeg', 'peg']:
                                newfile._name = generar_nombre("certificado_", newfile._name)
                            else:
                                return JsonResponse(
                                    {"result": True, "mensaje": u"Error, solo archivos .pdf,.jpg, .jpeg"})
                            vacunacion.certificado = newfile
                        vacunacion.save(request)
                        if vacunado:
                            datosdosis = request.POST.getlist('infoDosis[]')
                            if datosdosis:
                                c = 0
                                while c < len(datosdosis):
                                    dosis = VacunaCovidDosis(cabvacuna=vacunacion, numdosis=datosdosis[c],
                                                             fechadosis=datosdosis[c + 1])
                                    dosis.save(request)
                                    c += 2
                            else:
                                transaction.set_rollback(True)
                                return JsonResponse({"result": True, "mensaje": "Registre al menos una dosis."},
                                                    safe=False)

                        log(u'Adiciono Certificado de vacunación: %s' % vacunacion, request, "add")
                        return JsonResponse({"result": False, 'to': request.path}, safe=False)
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({"result": True, "mensaje": "Complete los datos requeridos."}, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        elif action == 'adddosisvacunaaspirante':
            try:
                with transaction.atomic():
                    vacunacion = VacunaCovid.objects.get(pk=int(request.POST['id']))
                    if 'recibiodosiscompleta' in request.POST:
                        recibiodosiscompleta = True if int(request.POST['recibiodosiscompleta']) == 1 else False
                        vacunacion.recibiodosiscompleta = recibiodosiscompleta
                    vacunacion.save(request)
                    VacunaCovidDosis.objects.filter(cabvacuna=vacunacion).delete()
                    datosdosis = request.POST.getlist('infoDosis[]')
                    if datosdosis:
                        c = 0
                        while c < len(datosdosis):
                            dosis = VacunaCovidDosis(cabvacuna=vacunacion, numdosis=datosdosis[c],
                                                     fechadosis=datosdosis[c + 1])
                            dosis.save(request)
                            c += 2
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({"result": True, "mensaje": "Registre al menos una dosis."}, safe=False)
                    log(u'Adiciono Certificado de vacunación: %s' % vacunacion, request, "add")
                    return JsonResponse({"result": False, 'to': request.path}, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        elif action == 'detalle_sustituto':
            try:
                data['personadatosfamiliares'] = personadatosfamiliares = PersonaDatosFamiliares.objects.get(
                    pk=int(request.POST['id']))
                template = get_template("th_hojavida/detalle_sustituto.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'data': json_content})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'subirfoto':
            try:
                f = PersonaSubirFotoForm(request.POST, request.FILES)

                if 'foto' in request.FILES:
                    arch = request.FILES['foto']
                    extension = arch._name.split('.')
                    tam = len(extension)
                    exte = extension[tam - 1]
                    if arch.size > 524288:
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"Error, el tamaño del archivo es mayor a 5 Kb."})
                    if not exte.lower() == 'jpg':
                        return JsonResponse({"result": "bad", "mensaje": u"Solo se permiten archivos .jpg"})

                if f.is_valid():
                    newfile = request.FILES['foto']
                    newfile._name = generar_nombre("foto_", newfile._name)

                    if not persona.tiene_foto():
                        fotopersona = FotoPersona(persona=persona,
                                                  foto=newfile)
                    else:
                        fotopersona = persona.fotopersona_set.all()[0]
                        fotopersona.foto = newfile

                    fotopersona.save(request)

                    log(u'Agregó foto: %s' % (persona), request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                msg = ex.__str__()
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos. Detalle: %s" % (msg)})

        if action == 'editfoto':
            try:
                foto = request.FILES['foto']
                max_tamano = 2 * 1024 * 1024  # 2 MB
                name_ = foto._name
                ext = name_[name_.rfind("."):]
                if not ext.lower() in ['.png', '.jpg', '.jpeg']:
                    raise NameError('Solo se permite archivos de formato .png, .jpg, .jpeg')

                if foto.size > max_tamano:
                    raise NameError('Archivo supera los 2 megas permitidos')
                # Asignar un nombre personalizado al archivo
                foto.name = unidecode(generar_nombre(f"foto_{persona.usuario}", foto._name))
                persona.foto(foto, request)
                persona = Persona.objects.get(id=persona.id)
                request.session['persona'] = persona
                log(u'Edito foto de perfil de ususario: %s' % foto, request, "edit")
                return JsonResponse({"result": False}, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': f'Error:{ex}'})

        elif action == 'addrevista':
            try:
                f = RevistaInvestigacionForm(request.POST, request.FILES)
                tiporegistro = int(request.POST['tiporegistro'])

                if f.is_valid():
                    if not RevistaInvestigacion.objects.filter(nombre=f.cleaned_data['nombrerevista'],
                                                               codigoissn=f.cleaned_data['codigoissn'],
                                                               tiporegistro=tiporegistro).exists():
                        revista = RevistaInvestigacion(nombre=f.cleaned_data['nombrerevista'],
                                                       codigoissn=f.cleaned_data['codigoissn'],
                                                       institucion=f.cleaned_data['institucion'],
                                                       tipo=f.cleaned_data['tipo'],
                                                       enlace=f.cleaned_data['enlacerevista'],
                                                       tiporegistro=tiporegistro,
                                                       borrador=True)
                        revista.save(request)

                        bases = f.cleaned_data['baseindexada']
                        for base in bases:
                            revistabase = RevistaInvestigacionBase(revista=revista,
                                                                   baseindexada_id=base.id,
                                                                   documentoindexacion=None
                                                                   )
                            revistabase.save(request)

                        log(u'Adiciono borrador revista/congreso de investigacion: %s [%s]' % (revista, revista.id),
                            request, "add")
                        return JsonResponse({"result": "ok", "id": revista.id})
                    else:
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"La Revista/Congreso ya ha sido ingresada anteriormente."})
                else:
                    raise NameError('Error')
            except Exception as ex:
                msg = ex.__str__()
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos. Detalle: %s" % (msg)})

        elif action == 'editrevista':
            try:
                f = RevistaInvestigacionHojaVidaForm(request.POST, request.FILES)

                revista = RevistaInvestigacion.objects.get(pk=request.POST['idrev'])
                if f.is_valid():
                    if not RevistaInvestigacion.objects.filter(nombre=f.cleaned_data['nombrerevista2'],
                                                               codigoissn=f.cleaned_data['codigoissn2']).exclude(
                        pk=request.POST['idrev']).exists():
                        revista.nombre = f.cleaned_data['nombrerevista2']
                        revista.codigoissn = f.cleaned_data['codigoissn2']
                        revista.institucion = f.cleaned_data['institucion2']
                        revista.enlace = f.cleaned_data['enlacerevista2']
                        revista.tipo = f.cleaned_data['tipo2']
                        revista.save(request)

                        bases = f.cleaned_data['baseindexada2']
                        basesnoseleccionadas = BaseIndexadaInvestigacion.objects.filter(
                            revistainvestigacionbase__revista=revista, revistainvestigacionbase__status=True).exclude(
                            pk__in=bases).order_by('nombre')

                        for base in bases:
                            if not RevistaInvestigacionBase.objects.filter(baseindexada=base, revista=revista).exists():
                                revistabase = RevistaInvestigacionBase(revista=revista,
                                                                       baseindexada_id=base.id,
                                                                       documentoindexacion=None
                                                                       )
                                revistabase.save(request)
                            elif RevistaInvestigacionBase.objects.filter(baseindexada=base, revista=revista,
                                                                         status=False).exists():
                                RevistaInvestigacionBase.objects.filter(revista=revista, baseindexada=base).update(
                                    status=True)

                        RevistaInvestigacionBase.objects.filter(revista=revista,
                                                                baseindexada__in=basesnoseleccionadas).update(
                            status=False)

                        log(u'Edito revista de investigacion: %s [%s]' % (revista, revista.id), request, "edit")
                        return JsonResponse({"result": "ok", "idrevista": revista.id})
                    else:
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"La Revista ya ha sido ingresada anteriormente."})
                else:
                    raise NameError('Error')
            except Exception as ex:
                msg = ex.__str__()
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos. Detalle: %s" % (msg)})

        elif action == 'addparticipantedocente':
            try:
                f = ParticipanteProfesorArticuloForm(request.POST)
                if f.is_valid():
                    if f.cleaned_data['profesor'] == 0:
                        return JsonResponse({"result": "bad", "mensaje": u"Seleccione un profesor."})

                    if ParticipanteSolicitudPublicacion.objects.filter(status=True, solicitud_id=request.POST['id'],
                                                                       profesor_id=f.cleaned_data['profesor']).exists():
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"El participante ya fue agregado anteriormente."})
                    else:
                        participante = ParticipanteSolicitudPublicacion(solicitud_id=request.POST['id'],
                                                                        profesor_id=f.cleaned_data['profesor'],
                                                                        tipo=f.cleaned_data['tipo'],
                                                                        tipoparticipanteins=f.cleaned_data[
                                                                            'tipoparticipanteins']
                                                                        )
                        participante.save(request)
                        log(u'Adiciono participante docentes en la solicitud: %s [%s]' % (
                            participante, participante.id), request, "add")
                        return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                msg = ex.__str__()
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos. Detalle: %s" % (msg)})

        elif action == 'addparticipanteadministrativo':
            try:
                f = ParticipanteAdministrativoArticuloForm(request.POST)
                if f.is_valid():
                    if f.cleaned_data['administrativo'] == 0:
                        return JsonResponse({"result": "bad", "mensaje": u"Seleccione un administrativo."})

                    if ParticipanteSolicitudPublicacion.objects.filter(status=True, solicitud_id=request.POST['id'],
                                                                       administrativo_id=f.cleaned_data[
                                                                           'administrativo']).exists():
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"El participante ya fue agregado anteriormente."})
                    else:
                        participante = ParticipanteSolicitudPublicacion(solicitud_id=request.POST['id'],
                                                                        administrativo_id=f.cleaned_data[
                                                                            'administrativo'],
                                                                        tipo=f.cleaned_data['tipo'],
                                                                        tipoparticipanteins=f.cleaned_data[
                                                                            'tipoparticipanteins']
                                                                        )
                        participante.save(request)
                        log(u'Adiciono participante administrativo en la solicitud: %s [%s]' % (
                            participante, participante.id), request, "add")
                        return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                msg = ex.__str__()
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos. Detalle: %s" % (msg)})

        elif action == 'addparticipanteinscripcion':
            try:
                f = ParticipanteInscripcionArticuloForm(request.POST)
                if f.is_valid():
                    if f.cleaned_data['inscripcion'] == 0:
                        return JsonResponse({"result": "bad", "mensaje": u"Seleccione un estudiante."})

                    if ParticipanteSolicitudPublicacion.objects.filter(status=True, solicitud_id=request.POST['id'],
                                                                       inscripcion_id=f.cleaned_data[
                                                                           'inscripcion']).exists():
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"El participante ya fue agregado anteriormente."})
                    else:
                        participante = ParticipanteSolicitudPublicacion(solicitud_id=request.POST['id'],
                                                                        inscripcion_id=f.cleaned_data['inscripcion'],
                                                                        tipo=f.cleaned_data['tipo'],
                                                                        tipoparticipanteins=f.cleaned_data[
                                                                            'tipoparticipanteins']
                                                                        )
                        participante.save(request)
                        log(u'Adiciono participante estudiante en la solicitud: %s [%s]' % (
                            participante, participante.id), request, "add")
                        return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                msg = ex.__str__()
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos. Detalle: %s" % (msg)})

        elif action == 'deleteparticipantepublicacion':
            try:
                participante = ParticipanteSolicitudPublicacion.objects.get(pk=request.POST['id'])
                participante.status = False
                participante.save(request)
                log(u'Elimino participante de publicación: %s [%s]' % (participante, participante.id), request, "del")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        elif action == 'updatetipoparticipante':
            try:
                # participantes = ParticipantesArticulos.objects.get(pk=request.POST['iditem'])
                participante = ParticipanteSolicitudPublicacion.objects.get(pk=request.POST['iditem'])
                idtipo = request.POST['idtipo']
                participante.tipoparticipanteins = idtipo
                participante.save(request)
                log(u'Actualizo tipo de participante en solicitud de publicación: %s [%s]' % (
                    participante, participante.id), request, "edit")
                return JsonResponse({'result': 'ok', 'idsolicitud': participante.solicitud.id})
            except Exception as ex:
                msg = ex.__str__()
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos. Detalle: %s" % (msg)})

        elif action == 'datosinstitucionales':
            try:
                persona = request.session['persona']
                f = DatosInstitucionalesPersonaForm(request.POST)
                if f.is_valid():
                    persona.telefonoextension = f.cleaned_data['extension']
                    persona.save(request)
                    log(u'Modifico extension: %s' % persona, request, "edit")
                    return JsonResponse({'result': 'ok'})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', 'mensaje': u'Error al guardar los datos'})

        elif action == 'datosnacimientoaspirante':
            try:
                persona = request.session['persona']
                f = DatosNacimientoAspiranteForm(request.POST)
                if f.is_valid():
                    persona.paisnacimiento = f.cleaned_data['paisnacimiento']
                    persona.provincianacimiento = f.cleaned_data['provincianacimiento']
                    persona.cantonnacimiento = f.cleaned_data['cantonnacimiento']
                    persona.parroquianacimiento = f.cleaned_data['parroquianacimiento']
                    persona.nacionalidad = f.cleaned_data['nacionalidad']
                    persona.save(request)
                    log(u'Modifico datos de nacimiento: %s' % persona, request, "edit")
                    return JsonResponse({'result': 'ok'})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', 'mensaje': u'Error al guardar los datos'})

        elif action == 'datosdomicilioaspirante':
            try:
                if 'archivocroquis' in request.FILES:
                    newfile = request.FILES['archivocroquis']
                    if newfile.size > 2194304:
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"Error, Tamaño de archivo Maximo permitido es de 2Mb"})

                if 'archivoplanillaluz' in request.FILES:
                    newfile = request.FILES['archivoplanillaluz']
                    if newfile.size > 2194304:
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"Error, Tamaño de archivo Maximo permitido es de 2Mb"})

                if 'serviciosbasico' in request.FILES:
                    newfile2 = request.FILES['archivoplanillaluz']
                    if newfile2.size > 2194304:
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"Error, Tamaño de archivo Maximo permitido es de 2Mb"})

                persona = request.session['persona']
                f = DatosDomicilioAdmisionForm(request.POST)
                if 'pais' in request.POST and request.POST['pais'] and int(request.POST['pais']) == 1:
                    if 'provincia' in request.POST and not request.POST['provincia']:
                        raise NameError('Debe ingresa una provincia')
                    if 'canton' in request.POST and not request.POST['canton']:
                        raise NameError('Debe ingresa una canton')
                    if 'parroquia' in request.POST and not request.POST['parroquia']:
                        raise NameError('Debe ingresa una parroquia')

                if f.is_valid():
                    newfile = None
                    persona.pais = f.cleaned_data['pais']
                    persona.provincia = f.cleaned_data['provincia']
                    persona.canton = f.cleaned_data['canton']
                    persona.sector = f.cleaned_data['sector']
                    persona.parroquia = f.cleaned_data['parroquia']
                    persona.direccion = f.cleaned_data['direccion']
                    persona.direccion2 = f.cleaned_data['direccion2']
                    persona.ciudadela = f.cleaned_data['ciudadela']
                    persona.num_direccion = f.cleaned_data['num_direccion']
                    persona.telefono_conv = f.cleaned_data['telefono_conv']
                    persona.telefono = f.cleaned_data['telefono']
                    persona.tipocelular = f.cleaned_data['tipocelular']
                    persona.referencia = f.cleaned_data['referencia']
                    persona.zona = int(f.cleaned_data['zona'])
                    persona.save(request)

                    docu = PersonaDocumentoPersonal.objects.get(status=True, persona=persona)
                    if 'archivocroquis' in request.FILES:
                        newfile = request.FILES['archivocroquis']
                        newfile._name = generar_nombre("croquis_", newfile._name)
                        persona.archivocroquis = newfile
                        persona.save(request)

                    if 'archivoplanillaluz' in request.FILES:
                        newfile = request.FILES['archivoplanillaluz']
                        newfile._name = generar_nombre("planilla_luz_", newfile._name)
                        persona.archivoplanillaluz = newfile
                        persona.save(request)

                    if 'serviciosbasico' in request.FILES:
                        newfile = request.FILES['serviciosbasico']
                        newfile._name = generar_nombre("serviciobasico_", newfile._name)
                        docu.serviciosbasico = newfile
                        docu.save(request)

                    log(u'Modifico datos de domicilio: %s' % persona, request, "edit")
                    return JsonResponse({'result': 'ok'})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', 'mensaje': u'Error al guardar los datos, %s' % ex})

        elif action == 'datosmedicos':
            try:
                persona = request.session['persona']
                f = DatosMedicosForm(request.POST)
                if f.is_valid():
                    datosextension = persona.datos_extension()
                    examenfisico = persona.datos_examen_fisico()
                    datosextension.carnetiess = f.cleaned_data['carnetiess']
                    # if perfilprincipal.es_estudiante():
                    personaexamenfisico = \
                        PersonaExamenFisico.objects.filter(personafichamedica__personaextension=datosextension)[0]
                    personaexamenfisico.peso = f.cleaned_data['peso']
                    personaexamenfisico.talla = f.cleaned_data['talla']
                    personaexamenfisico.save(request)
                    persona.sangre = f.cleaned_data['sangre']
                    persona.save(request)
                    datosextension.save(request)

                    if 'archivotiposangre' in request.FILES:
                        newfile = request.FILES['archivotiposangre']
                        newfile._name = generar_nombre("tiposangre", newfile._name)

                        documento = persona.documentos_personales()
                        if documento is None:
                            documento = PersonaDocumentoPersonal(persona=persona,
                                                                 tiposangre=newfile,
                                                                 estadotiposangre=1
                                                                 )
                        else:
                            documento.tiposangre = newfile
                            documento.estadotiposangre = 1

                        documento.save(request)

                    log(u'Modifico datos de medicos basicos: %s' % persona, request, "edit")
                    return JsonResponse({'result': 'ok'})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', 'mensaje': u'Error al guardar los datos'})

        elif action == 'contactoemergencia':
            try:
                persona = request.session['persona']
                f = ContactoEmergenciaForm(request.POST)
                if f.is_valid():
                    datosextension = persona.datos_extension()
                    datosextension.contactoemergencia = f.cleaned_data['contactoemergencia']
                    datosextension.telefonoemergencia = f.cleaned_data['telefonoemergencia']
                    datosextension.parentescoemergencia = f.cleaned_data['parentescoemergencia']
                    datosextension.save(request)
                    log(u'Modifico datos de contacto de emergencia: %s' % persona, request, "edit")
                    return JsonResponse({'result': 'ok'})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', 'mensaje': u'Error al guardar los datos'})

        elif action == 'contactoemergenciaaspirante':
            try:
                persona = request.session['persona']
                f = ContactoEmergenciaAspiranteForm(request.POST)
                if f.is_valid():
                    datosextension = persona.datos_extension()
                    datosextension.contactoemergencia = f.cleaned_data['contactoemergencia']
                    datosextension.telefonoemergencia = f.cleaned_data['telefonoemergencia']
                    datosextension.parentescoemergencia = f.cleaned_data['parentescoemergencia']
                    datosextension.save(request)
                    log(u'Modifico datos de contacto de emergencia: %s' % persona, request, "edit")
                    return JsonResponse({'result': 'ok'})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', 'mensaje': u'Error al guardar los datos'})

        elif action == 'etnia':
            try:
                persona = request.session['persona']
                if 'archivoraza' in request.FILES:
                    arch = request.FILES['archivoraza']
                    extension = arch._name.split('.')
                    tam = len(extension)
                    exte = extension[tam - 1]
                    if arch.size > 4194304:
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"Error, el tamaño del archivo es mayor a 4 Mb."})
                    if not exte.lower() in ['pdf']:
                        return JsonResponse({"result": "bad", "mensaje": u"Solo se permiten archivos .pdf"})

                f = EtniaForm(request.POST, request.FILES)
                if f.is_valid():
                    perfil = persona.mi_perfil()
                    perfil.raza = f.cleaned_data['raza']
                    perfil.nacionalidadindigena = f.cleaned_data['nacionalidadindigena']

                    if 'archivoraza' in request.FILES:
                        arch._name = generar_nombre("archivoraza", arch._name)
                        perfil.archivoraza = arch
                        perfil.estadoarchivoraza = 1
                    else:
                        if perfil.raza.id != 1:
                            perfil.archivoraza = None
                            perfil.estadoarchivoraza = None

                    perfil.save(request)
                    log(u'Modifico etnia: %s' % persona, request, "edit")
                    return JsonResponse({'result': 'ok'})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', 'mensaje': u'Error al guardar los datos'})

        elif action == 'etniaaspirante':
            try:
                persona = request.session['persona']
                if 'archivoraza' in request.FILES:
                    arch = request.FILES['archivoraza']
                    extension = arch._name.split('.')
                    tam = len(extension)
                    exte = extension[tam - 1]
                    if arch.size > 4194304:
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"Error, el tamaño del archivo es mayor a 4 Mb."})
                    if not exte.lower() in ['pdf']:
                        return JsonResponse({"result": "bad", "mensaje": u"Solo se permiten archivos .pdf"})

                f = EtniaAspiranteForm(request.POST, request.FILES)
                if f.is_valid():
                    perfil = persona.mi_perfil()
                    perfil.raza = f.cleaned_data['raza']
                    perfil.nacionalidadindigena = f.cleaned_data['nacionalidadindigena']

                    if 'archivoraza' in request.FILES:
                        arch._name = generar_nombre("archivoraza", arch._name)
                        perfil.archivoraza = arch
                        perfil.estadoarchivoraza = 1
                    else:
                        if perfil.raza.id != 1:
                            perfil.archivoraza = None
                            perfil.estadoarchivoraza = None

                    perfil.save(request)
                    log(u'Modifico etnia: %s' % persona, request, "edit")
                    return JsonResponse({'result': 'ok'})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', 'mensaje': u'Error al guardar los datos'})

        elif action == 'discapacidadaspirante':
            try:
                if 'archivo' in request.FILES:
                    arch = request.FILES['archivo']
                    extension = arch._name.split('.')
                    tam = len(extension)
                    exte = extension[tam - 1]
                    if arch.size > 4194304:
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"Error, el tamaño del archivo es mayor a 4 Mb."})
                    if not exte.lower() in ['pdf']:
                        return JsonResponse({"result": "bad", "mensaje": u"Solo se permiten archivos .pdf"})
                if 'archivovaloracion' in request.FILES:
                    arch = request.FILES['archivovaloracion']
                    extension = arch._name.split('.')
                    tam = len(extension)
                    exte = extension[tam - 1]
                    if arch.size > 4194304:
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"Error, el tamaño del archivo es mayor a 4 Mb."})
                    if not exte.lower() in ['pdf']:
                        return JsonResponse({"result": "bad", "mensaje": u"Solo se permiten archivos .pdf"})

                persona = request.session['persona']
                f = DiscapacidadForm(request.POST, request.FILES)
                if f.is_valid():
                    newfile = None
                    if not f.cleaned_data['tienediscapacidad'] and f.cleaned_data['tienediscapacidadmultiple']:
                        raise NameError('No puede marcar discapacidad multiple sin marcar que tiene discapacidad')
                    if not f.cleaned_data['tipodiscapacidad'] and f.cleaned_data['tipodiscapacidadmultiple']:
                        raise NameError(
                            'No puede elegir discapacidades multiples sin elegir una discapacidad principal')
                    if f.cleaned_data['tienediscapacidadmultiple'] and not f.cleaned_data['tipodiscapacidadmultiple']:
                        raise NameError('Debe elegir una o más discapacidades multiples')
                    perfil = persona.mi_perfil()
                    perfil.tienediscapacidad = f.cleaned_data['tienediscapacidad']
                    perfil.tipodiscapacidad = f.cleaned_data['tipodiscapacidad']
                    perfil.porcientodiscapacidad = f.cleaned_data['porcientodiscapacidad'] if f.cleaned_data[
                        'porcientodiscapacidad'] else 0
                    perfil.carnetdiscapacidad = f.cleaned_data['carnetdiscapacidad']
                    perfil.institucionvalida = f.cleaned_data['institucionvalida']
                    perfil.tienediscapacidadmultiple = f.cleaned_data['tienediscapacidadmultiple']
                    perfil.grado = f.cleaned_data['grado'] if f.cleaned_data['grado'] else 0

                    if not f.cleaned_data['tienediscapacidad']:
                        perfil.archivo = None
                        perfil.estadoarchivodiscapacidad = None
                        perfil.archivovaloracion = None
                    if 'archivo' in request.FILES:
                        newfile = request.FILES['archivo']
                        newfile._name = generar_nombre("archivosdiscapacidad_", newfile._name)
                        perfil.archivo = newfile
                        perfil.estadoarchivodiscapacidad = 1
                    if 'archivovaloracion' in request.FILES:
                        newfile = request.FILES['archivovaloracion']
                        newfile._name = generar_nombre("archivovaloracionmedica_", newfile._name)
                        perfil.archivovaloracion = newfile
                    perfil.save(request)
                    perfil.tipodiscapacidadmultiple.clear()
                    perfil.subtipodiscapacidad.clear()
                    if f.cleaned_data['tienediscapacidadmultiple']:
                        tipos = request.POST.getlist('tipodiscapacidadmultiple')
                        for tipo in tipos:
                            perfil.tipodiscapacidadmultiple.add(tipo)
                    if f.cleaned_data['tipodiscapacidad'] and f.cleaned_data['tienediscapacidad']:
                        subtipos = request.POST.getlist('subtipodiscapacidad')
                        for subtipo in subtipos:
                            perfil.subtipodiscapacidad.add(subtipo)
                    log(u'Modifico tipo de discapacidad: %s' % persona, request, "edit")
                    return JsonResponse({'result': 'ok'})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', 'mensaje': u'Error al guardar los datos'})

        elif action == 'buscarsubtipo':
            try:
                tipo = request.POST['id']
                subtipo = SubTipoDiscapacidad.objects.filter(discapacidad_id=tipo, status=True)
                return JsonResponse(
                    {'result': 'ok', 'lista': list(subtipo.values_list('id', 'nombre')) if subtipo.exists() else []})
            except Exception as ex:
                print(ex)
                return JsonResponse({'result': 'bad', 'mensaje': u'Error al guardar los datos'})

        elif action == 'addfamiliaraspirante':
            try:
                persona = request.session['persona']
                f = FamiliarAspiranteForm(request.POST)
                if f.is_valid():
                    cedula = f.cleaned_data['identificacion'].strip()
                    if persona.personadatosfamiliares_set.filter(identificacion=f.cleaned_data['identificacion']).exists():
                        raise NameError('El familiar se encuentra registrado.')
                    resp = validarcedula(cedula)
                    if resp != 'Ok':
                        raise NameError(resp)
                    nombres = f"{f.cleaned_data['apellido1']} {f.cleaned_data['apellido2']} {f.cleaned_data['nombre']}"
                    familiar = PersonaDatosFamiliares(persona=persona,
                                                      identificacion=cedula,
                                                      nombre=nombres,
                                                      fallecido=f.cleaned_data['fallecido'],
                                                      nacimiento=f.cleaned_data['nacimiento'],
                                                      parentesco=f.cleaned_data['parentesco'],
                                                      tienediscapacidad=f.cleaned_data['tienediscapacidad'],
                                                      telefono=f.cleaned_data['telefono'],
                                                      telefono_conv=f.cleaned_data['telefono_conv'],
                                                      niveltitulacion=f.cleaned_data['niveltitulacion'],
                                                      ingresomensual=f.cleaned_data['ingresomensual'],
                                                      formatrabajo=f.cleaned_data['formatrabajo'],
                                                      trabajo=f.cleaned_data['trabajo'],
                                                      convive=f.cleaned_data['convive'],
                                                      sustentohogar=f.cleaned_data['sustentohogar'],
                                                      rangoedad=f.cleaned_data['rangoedad'],
                                                      essustituto=f.cleaned_data['essustituto'],
                                                      autorizadoministerio=f.cleaned_data['autorizadoministerio'],
                                                      tipodiscapacidad=f.cleaned_data['tipodiscapacidad'],
                                                      porcientodiscapacidad=f.cleaned_data['porcientodiscapacidad'],
                                                      carnetdiscapacidad=f.cleaned_data['carnetdiscapacidad'],
                                                      institucionvalida=f.cleaned_data['institucionvalida'],
                                                      tipoinstitucionlaboral=f.cleaned_data['tipoinstitucionlaboral'],
                                                      negocio=f.cleaned_data['negocio'],
                                                      tienenegocio=f.cleaned_data['tienenegocio'])
                    familiar.save(request)
                    if 'cedulaidentidad' in request.FILES:
                        newfile = request.FILES['cedulaidentidad']
                        newfile._name = generar_nombre("cedulaidentidad_", newfile._name)
                        familiar.cedulaidentidad = newfile
                        familiar.save(request)
                    if 'ceduladiscapacidad' in request.FILES:
                        newfile = request.FILES['ceduladiscapacidad']
                        newfile._name = generar_nombre("ceduladiscapacidad_", newfile._name)
                        familiar.ceduladiscapacidad = newfile
                        familiar.save(request)
                    if 'archivoautorizado' in request.FILES:
                        newfile = request.FILES['archivoautorizado']
                        newfile._name = generar_nombre("archivoautorizado_", newfile._name)
                        familiar.archivoautorizado = newfile
                        familiar.save(request)

                    pers = Persona.objects.filter(
                        Q(pasaporte=cedula) | Q(cedula=cedula) | Q(pasaporte=('VS' + cedula)) | Q(cedula=cedula[2:]),
                        status=True).first()
                    if not pers:
                        pers = Persona(cedula=f.cleaned_data['identificacion'],
                                       nombres=f.cleaned_data['nombre'],
                                       apellido1=f.cleaned_data['apellido1'],
                                       apellido2=f.cleaned_data['apellido2'],
                                       nacimiento=f.cleaned_data['nacimiento'],
                                       telefono=f.cleaned_data['telefono'],
                                       telefono_conv=f.cleaned_data['telefono_conv'],
                                       )
                        pers.save(request)
                        log(u'Adiciono persona: %s' % persona, request, "add")
                    elif len(pers.mis_perfilesusuarios()) == 1 and pers.tiene_usuario_externo():
                        pers.cedula = f.cleaned_data['identificacion']
                        pers.nombres = f.cleaned_data['nombre']
                        pers.apellido1 = f.cleaned_data['apellido1']
                        pers.apellido2 = f.cleaned_data['apellido2']
                        pers.nacimiento = f.cleaned_data['nacimiento']
                        pers.telefono = f.cleaned_data['telefono']
                        pers.telefono_conv = f.cleaned_data['telefono_conv']
                        pers.save(request)
                        log(u'Edito familiar de usuario: %s' % pers, request, "edit")
                    if not pers.tiene_perfil():
                        externo = Externo(persona=pers)
                        externo.save(request)
                        log(u'Adiciono externo: %s' % pers, request, "add")
                        perfil = PerfilUsuario(persona=pers, externo=externo)
                        perfil.save(request)
                        log(u'Adiciono perfil de usuario: %s' % perfil, request, "add")
                    if f.cleaned_data['parentesco'].id in [14, 11] and not persona.apellido1 in [pers.apellido1,
                                                                                                 pers.apellido2]:
                        familiar.aprobado = False
                    familiar.personafamiliar = pers
                    familiar.save(request)
                    log(u'Adiciono familiar: %s' % familiar, request, "add")
                    return JsonResponse({'result': 'ok'})
                else:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()],
                                         "mensaje": "Error en el formulario"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', 'mensaje': u'Error al guardar los datos'})

        elif action == 'editfamiliaraspirante':
            try:
                persona = request.session['persona']
                f = FamiliarAspiranteForm(request.POST)
                f.edit()
                if f.is_valid():
                    familiar = PersonaDatosFamiliares.objects.get(pk=int(request.POST['id']))
                    cedula = f.cleaned_data['identificacion'].strip()
                    if persona.personadatosfamiliares_set.filter(identificacion=cedula).exclude(
                            id=familiar.id).exists():
                        raise NameError(u'El familiar se encuentra registrado.')
                    resp = validarcedula(cedula)
                    if resp != 'Ok':
                        raise NameError(resp)
                    nombres = f"{f.cleaned_data['apellido1']} {f.cleaned_data['apellido2']} {f.cleaned_data['nombre']}"
                    pers = Persona.objects.filter(cedula=cedula, status=True).first()
                    if len(pers.mis_perfilesusuarios()) == 1 and pers.tiene_usuario_externo():
                        familiar.identificacion = cedula
                        familiar.nombre = nombres
                        familiar.nacimiento = f.cleaned_data['nacimiento']
                        familiar.telefono = f.cleaned_data['telefono']
                        familiar.telefono_conv = f.cleaned_data['telefono_conv']
                    familiar.fallecido = f.cleaned_data['fallecido']
                    familiar.parentesco = f.cleaned_data['parentesco']
                    familiar.tienediscapacidad = f.cleaned_data['tienediscapacidad']
                    familiar.trabajo = f.cleaned_data['trabajo']
                    familiar.niveltitulacion = f.cleaned_data['niveltitulacion']
                    familiar.ingresomensual = f.cleaned_data['ingresomensual']
                    familiar.formatrabajo = f.cleaned_data['formatrabajo']
                    familiar.convive = f.cleaned_data['convive']
                    familiar.sustentohogar = f.cleaned_data['sustentohogar']
                    familiar.rangoedad = f.cleaned_data['rangoedad']
                    familiar.tienenegocio = f.cleaned_data['tienenegocio']
                    if f.cleaned_data['tienediscapacidad']:
                        familiar.essustituto = f.cleaned_data['essustituto']
                        familiar.autorizadoministerio = f.cleaned_data['autorizadoministerio']
                        familiar.tipodiscapacidad = f.cleaned_data['tipodiscapacidad']
                        familiar.porcientodiscapacidad = f.cleaned_data['porcientodiscapacidad']
                        familiar.carnetdiscapacidad = f.cleaned_data['carnetdiscapacidad']
                        familiar.institucionvalida = f.cleaned_data['institucionvalida']
                    else:
                        familiar.essustituto = False
                        familiar.autorizadoministerio = False
                        familiar.tipodiscapacidad = None
                        familiar.porcientodiscapacidad = None
                        familiar.carnetdiscapacidad = ''
                        familiar.institucionvalida = None
                    familiar.negocio = ''
                    familiar.tipoinstitucionlaboral = f.cleaned_data['tipoinstitucionlaboral']
                    if familiar.tienenegocio:
                        familiar.negocio = f.cleaned_data['negocio']
                    familiar.save(request)
                    if 'cedulaidentidad' in request.FILES:
                        newfile = request.FILES['cedulaidentidad']
                        newfile._name = generar_nombre("cedulaidentidad_", newfile._name)
                        familiar.cedulaidentidad = newfile
                        familiar.save(request)
                    if 'ceduladiscapacidad' in request.FILES:
                        newfile = request.FILES['ceduladiscapacidad']
                        newfile._name = generar_nombre("ceduladiscapacidad_", newfile._name)
                        familiar.ceduladiscapacidad = newfile
                        familiar.save(request)
                    if f.cleaned_data['tienediscapacidad']:
                        if 'archivoautorizado' in request.FILES:
                            newfile = request.FILES['archivoautorizado']
                            newfile._name = generar_nombre("archivoautorizado_", newfile._name)
                            familiar.archivoautorizado = newfile
                            familiar.save(request)
                    else:
                        familiar.archivoautorizado = None
                        familiar.save(request)
                    if not pers:
                        pers = Persona(cedula=f.cleaned_data['identificacion'],
                                       nombres=f.cleaned_data['nombre'],
                                       apellido1=f.cleaned_data['apellido1'],
                                       apellido2=f.cleaned_data['apellido2'],
                                       nacimiento=f.cleaned_data['nacimiento'],
                                       telefono=f.cleaned_data['telefono'],
                                       telefono_conv=f.cleaned_data['telefono_conv'],
                                       )
                        pers.save(request)
                        log(u'Adiciono persona: %s' % persona, request, "add")
                    elif len(pers.mis_perfilesusuarios()) == 1 and pers.tiene_usuario_externo():
                        pers.cedula = f.cleaned_data['identificacion']
                        pers.nombres = f.cleaned_data['nombre']
                        pers.apellido1 = f.cleaned_data['apellido1']
                        pers.apellido2 = f.cleaned_data['apellido2']
                        pers.nacimiento = f.cleaned_data['nacimiento']
                        pers.telefono = f.cleaned_data['telefono']
                        pers.telefono_conv = f.cleaned_data['telefono_conv']
                        pers.save(request)
                        log(u'Edito familiar de usuario: %s' % pers, request, "edit")
                    if not pers.tiene_perfil():
                        externo = Externo(persona=pers)
                        externo.save(request)
                        log(u'Adiciono externo: %s' % pers, request, "add")
                        perfil = PerfilUsuario(persona=pers, externo=externo)
                        perfil.save(request)
                        log(u'Adiciono perfil de usuario: %s' % perfil, request, "add")
                    if f.cleaned_data['parentesco'].id in [14, 11] and not persona.apellido1 in [pers.apellido1,
                                                                                                 pers.apellido2]:
                        familiar.aprobado = False
                    familiar.personafamiliar = pers
                    familiar.save(request)
                    log(u'Modifico familiar: %s' % persona, request, "edit")
                    return JsonResponse({'result': 'ok'})
                else:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()],
                                         "mensaje": "Error en el formulario"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', 'mensaje': u'Error al guardar los datos'})

        elif action == 'respondehijo':
            try:
                persona = request.session['persona']
                datosext = persona.datos_extension()
                if 'valor' in request.POST:
                    datosext.tienehijos = 2
                else:
                    datosext.tienehijos = 3
                datosext.save()
                log(u'Contestó encuesta de hijos: %s' % persona, request, "edit")
                return JsonResponse({'result': 'ok'})
                # return JsonResponse({'result': 'ok'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', 'mensaje': u'Error al eliminar los datos'})

        elif action == 'delfamiliaraspirante':
            try:
                persona = request.session['persona']
                familiar = PersonaDatosFamiliares.objects.get(pk=int(request.POST['id']))
                familiar.delete()
                log(u'Elimino familiar: %s' % persona, request, "del")
                return JsonResponse({'result': True, 'error': False})
                # return JsonResponse({'result': 'ok'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', 'mensaje': u'Error al eliminar los datos'})

        elif action == 'delhistorialaspirante':
            try:
                registro = PersonaAportacionHistorialLaboral.objects.get(pk=int(request.POST['id']))
                registro.delete()
                log(u'Elimino familiar: %s' % registro, request, "del")
                return JsonResponse({'result': True, 'error': False})
                # return JsonResponse({'result': 'ok'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', 'mensaje': u'Error al eliminar los datos'})

        elif action == 'adddeclaracion':
            try:
                newfile = None
                if 'archivo' in request.FILES:
                    newfile = request.FILES['archivo']
                    if newfile.size > 4194304:
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"Error, Tamaño de archivo Maximo permitido es de 4Mb"})
                persona = request.session['persona']
                f = DeclaracionBienForm(request.POST, request.FILES)
                if f.is_valid():
                    if 'archivo' not in request.FILES:
                        return JsonResponse({'result': 'bad', 'mensaje': u'Debe subir la declaración en formato pdf'})

                    newfile = request.FILES['archivo']
                    newfile._name = generar_nombre("declaracion_", newfile._name)
                    fechaproxima = None
                    if int(f.cleaned_data['tipodeclaracion']) == 3:
                        fechaproxima = f.cleaned_data['fecha'] + timedelta(days=730)  # sumar dos años

                    distributivo = None
                    departamento = None
                    tipodeclaracion = f.cleaned_data['tipodeclaracion']

                    if f.cleaned_data['cargosvigentes']:
                        distributivo = DistributivoPersona.objects.filter(persona=persona, status=True, denominacionpuesto=f.cleaned_data['cargo']).first()
                        departamento = distributivo.unidadorganica
                    else:
                        if tipodeclaracion == '1':
                            if not f.cleaned_data['departamento']:
                                return JsonResponse({'result': 'bad', 'mensaje': u'Debe seleccionar un departamento'})
                            departamento = f.cleaned_data['departamento']
                        else:
                            distributivo = DistributivoPersonaHistorial.objects.filter(persona=persona, status=True, denominacionpuesto=f.cleaned_data['cargo']).first()
                            departamento = distributivo.unidadorganica

                    if DeclaracionBienes.objects.filter(persona=persona,
                                                        denominacionpuesto=f.cleaned_data['cargo'],
                                                        departamento=departamento,
                                                        fechaperiodoinicio=f.cleaned_data['fechaperiodoinicio'],
                                                        tipodeclaracion=tipodeclaracion,
                                                        status=True).exists():
                        return JsonResponse({'result': 'bad', 'mensaje': u'Ya tiene una declaración con estos datos'})

                    declaracion = DeclaracionBienes(persona=persona,
                                                    # numero=f.cleaned_data['numero'],
                                                    # provincia=f.cleaned_data['provincia'],
                                                    # canton=f.cleaned_data['canton'],
                                                    # parroquia=f.cleaned_data['parroquia'],
                                                    fecha=f.cleaned_data['fecha'],
                                                    tipodeclaracion=tipodeclaracion,
                                                    codigobarra=f.cleaned_data['codigobarra'],
                                                    fechaproximoregistro=fechaproxima,
                                                    denominacionpuesto=f.cleaned_data['cargo'],
                                                    departamento=departamento,
                                                    fechaperiodoinicio=f.cleaned_data['fechaperiodoinicio'],
                                                    archivo=newfile)
                    declaracion.save(request)
                    if newfile is not None:
                        if DistributivoPersona.objects.filter(status=True, persona=persona):
                            lista = ['talento_humano@unemi.edu.ec', 'portizg@unemi.edu.ec']
                            asunto = "Ingresaron nueva declaración (archivo)"
                            send_html_mail(asunto, "emails/ingreso_archivos_tthh.html",
                                           {'asunto': asunto, 'd': declaracion.persona.nombre_completo_inverso(),
                                            'fecha': datetime.now().date(), 'escenario': 'declaración'}, lista, [],
                                           cuenta=CUENTAS_CORREOS[1][1])
                        log(u'Adiciono declaracion: %s' % persona, request, "add")
                    return JsonResponse({'result': False, 'mensaje': 'Guardado con éxito'})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', 'mensaje': u'Error al guardar los datos'})


        elif action == 'editdeclaracion':
            try:
                declaracion = DeclaracionBienes.objects.get(pk=encrypt_id(request.POST['id']))
                newfile = None
                if 'archivo' in request.FILES:
                    newfile = request.FILES['archivo']
                    if newfile.size > 4194304:
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"Error, Tamaño de archivo Maximo permitido es de 4Mb"})
                persona = request.session['persona']
                f = DeclaracionBienForm(request.POST, request.FILES)
                if f.is_valid():
                    if 'archivo' in request.FILES:
                        newfile = request.FILES['archivo']
                        newfile._name = generar_nombre("declaracion_", newfile._name)

                    # newfile = request.FILES['archivo']
                    # newfile._name = generar_nombre("declaracion_", newfile._name)
                    fechaproxima = None
                    if int(f.cleaned_data['tipodeclaracion']) == 3:
                        fechaproxima = f.cleaned_data['fecha'] + timedelta(days=730)  # sumar dos años

                    distributivo = None
                    departamento = None
                    tipodeclaracion = f.cleaned_data['tipodeclaracion']

                    if f.cleaned_data['cargosvigentes']:
                        distributivo = DistributivoPersona.objects.filter(persona=persona, status=True,
                                                                          denominacionpuesto=f.cleaned_data[
                                                                              'cargo']).first()
                        departamento = distributivo.unidadorganica
                    else:
                        if tipodeclaracion == '1':
                            if not f.cleaned_data['departamento']:
                                return JsonResponse({'result': 'bad', 'mensaje': u'Debe seleccionar un departamento'})
                            departamento = f.cleaned_data['departamento']
                        else:
                            distributivo = DistributivoPersonaHistorial.objects.filter(persona=persona, status=True,
                                                                                       denominacionpuesto=
                                                                                       f.cleaned_data['cargo']).first()
                            departamento = distributivo.unidadorganica

                    if DeclaracionBienes.objects.filter(persona=persona,
                                                        denominacionpuesto=f.cleaned_data['cargo'],
                                                        departamento=departamento,
                                                        fechaperiodoinicio=f.cleaned_data['fechaperiodoinicio'],
                                                        tipodeclaracion=tipodeclaracion,
                                                        status=True).exclude(pk=declaracion.pk).exists():
                        return JsonResponse({'result': 'bad', 'mensaje': u'Ya tiene una declaración con estos datos'})

                    declaracion.fecha = f.cleaned_data['fecha']
                    declaracion.tipodeclaracion = tipodeclaracion
                    declaracion.codigobarra = f.cleaned_data['codigobarra']
                    declaracion.fechaperiodoinicio = f.cleaned_data['fechaperiodoinicio']
                    declaracion.fechaproximoregistro = fechaproxima
                    declaracion.denominacionpuesto = f.cleaned_data['cargo']
                    declaracion.departamento = departamento

                    if newfile is not None:
                        declaracion.archivo = newfile

                    declaracion.save(request)

                    if DistributivoPersona.objects.filter(status=True, persona=persona):
                        lista = ['talento_humano@unemi.edu.ec', 'portizg@unemi.edu.ec']
                        asunto = "Se modificó una declaración patrimonial"
                        send_html_mail(asunto, "emails/ingreso_archivos_tthh.html",
                                       {'asunto': asunto, 'd': declaracion.persona.nombre_completo_inverso(),
                                        'fecha': datetime.now().date(), 'escenario': 'declaración'}, lista, [],
                                       cuenta=CUENTAS_CORREOS[1][1])
                    log(u'Modifico declaracion: %s' % persona, request, "edit")

                    return JsonResponse({'result': False, 'mensaje': 'Guardado con éxito'})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', 'mensaje': u'Error al guardar los datos'})

        elif action == 'deldeclaracion':
            try:
                persona = request.session['persona']
                declaracion = DeclaracionBienes.objects.get(pk=encrypt_id(request.POST['id']))
                if declaracion.verificado:
                    return JsonResponse({'result': 'bad', 'mensaje': u'No se puede elminar la declaración'})
                declaracion.status = False
                declaracion.save()
                log(u'Elimino declaracion: %s' % persona, request, "del")
                return JsonResponse({'error': False, 'message': 'Eliminado con éxito'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'error': True, 'message': 'Error al eliminar los datos'})

        elif action == 'deldeclaracionaspirante':
            try:
                persona = request.session['persona']
                declaracion = DeclaracionBienes.objects.get(pk=int(request.POST['id']))
                if declaracion.verificado:
                    return JsonResponse({'result': 'bad', 'mensaje': u'No se puede elminar la declaración'})
                declaracion.delete()
                log(u'Elimino declaracion: %s' % persona, request, "del")
                return JsonResponse({'result': True, 'error': False})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', 'mensaje': u'Error al eliminar los datos'})

        elif action == 'delvacunacionaspirante':
            try:
                persona = request.session['persona']
                vacuna = VacunaCovid.objects.get(pk=int(request.POST['id']))
                vacuna.status = False
                vacuna.save(request)
                log(u'Elimino Evidencia de Vacunación: %s' % vacuna, request, "del")
                return JsonResponse({'result': True, 'error': False})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', 'mensaje': u'Error al eliminar los datos'})

        elif action == 'addtitulacionaspirante':
            try:
                persona = request.session['persona']
                f = TitulacionPersonaAdmisionForm(request.POST, request.FILES)
                if 'archivo' in request.FILES:
                    arch = request.FILES['archivo']
                    extencion = arch._name.split('.')
                    exte = extencion[1]
                    if arch.size > 4194304:
                        return JsonResponse({"result": "bad", "mensaje": u"Error, archivo mayor a 4 Mb."})
                    if not exte == 'pdf' and not exte == 'png' and not exte == 'jpg' and not exte == 'jpeg' and not exte == 'jpg':
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"Archivo Título solo en formato .pdf, jpg, jpeg, png"})
                if 'registroarchivo' in request.FILES:
                    registroarchivo = request.FILES['registroarchivo']
                    extencion1 = registroarchivo._name.split('.')
                    exte1 = extencion1[1]
                    if registroarchivo.size > 4194304:
                        return JsonResponse({"result": "bad", "mensaje": u"Error, archivo mayor a 4 Mb."})
                    if not exte1 == 'pdf' and not exte1 == 'png' and not exte1 == 'jpg' and not exte1 == 'jpeg' and not exte1 == 'jpg':
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"Archivo SENESCYT solo en formato .pdf, jpg, jpeg, png"})
                if f.is_valid():
                    titulacion = Titulacion(persona=persona,
                                            titulo=f.cleaned_data['titulo'],
                                            areatitulo=f.cleaned_data['areatitulo'],
                                            fechainicio=f.cleaned_data['fechainicio'],
                                            fechaobtencion=f.cleaned_data['fechaobtencion'],
                                            fechaegresado=f.cleaned_data['fechaegresado'],
                                            registro=f.cleaned_data['registro'],
                                            # areaconocimiento=f.cleaned_data['areaconocimiento'],
                                            # subareaconocimiento=f.cleaned_data['subareaconocimiento'],
                                            # subareaespecificaconocimiento=f.cleaned_data['subareaespecificaconocimiento'],
                                            pais=f.cleaned_data['pais'],
                                            provincia=f.cleaned_data['provincia'],
                                            canton=f.cleaned_data['canton'],
                                            parroquia=f.cleaned_data['parroquia'],
                                            educacionsuperior=f.cleaned_data['educacionsuperior'],
                                            institucion=f.cleaned_data['institucion'],
                                            colegio=f.cleaned_data['colegio'],
                                            anios=f.cleaned_data['anios'],
                                            semestres=f.cleaned_data['semestres'],
                                            cursando=f.cleaned_data['cursando'],
                                            aplicobeca=f.cleaned_data['aplicobeca'],
                                            tipobeca=f.cleaned_data['tipobeca'] if f.cleaned_data[
                                                'aplicobeca'] else None,
                                            financiamientobeca=f.cleaned_data['financiamientobeca'] if f.cleaned_data[
                                                'aplicobeca'] else None,
                                            valorbeca=f.cleaned_data['valorbeca'] if f.cleaned_data[
                                                'aplicobeca'] else 0)
                    titulacion.save(request)
                    if not titulacion.cursando:
                        if 'archivo' in request.FILES:
                            newfile = request.FILES['archivo']
                            if newfile:
                                newfile._name = generar_nombre("titulacion_", newfile._name)
                                titulacion.archivo = newfile
                                titulacion.save(request)
                            if DistributivoPersona.objects.filter(status=True, persona=persona):
                                lista = ['talento_humano@unemi.edu.ec', 'portizg@unemi.edu.ec']
                                asunto = "Ingresaron título académico (archivo)"
                                send_html_mail(asunto, "emails/ingreso_archivos_tthh.html",
                                               {'asunto': asunto, 'd': titulacion.persona.nombre_completo_inverso(),
                                                'fecha': datetime.now().date(), 'escenario': 'titulación - SENECYT'},
                                               lista, [], cuenta=CUENTAS_CORREOS[1][1])

                        if 'registroarchivo' in request.FILES:
                            newfile2 = request.FILES['registroarchivo']
                            if newfile2:
                                newfile2._name = generar_nombre("archivosenecyt_", newfile2._name)
                                titulacion.registroarchivo = newfile2
                                titulacion.save(request)
                            if DistributivoPersona.objects.filter(status=True, persona=persona):
                                lista = ['talento_humano@unemi.edu.ec', 'portizg@unemi.edu.ec']
                                asunto = "Ingresaron título académico (archivo SENESCYT)"
                                send_html_mail(asunto, "emails/ingreso_archivos_tthh.html",
                                               {'asunto': asunto, 'd': titulacion.persona.nombre_completo_inverso(),
                                                'fecha': datetime.now().date(),
                                                'escenario': 'titulación - título academico'}, lista, [],
                                               cuenta=CUENTAS_CORREOS[1][1])
                    campotitulo = None
                    if CamposTitulosPostulacion.objects.filter(status=True, titulo=f.cleaned_data['titulo']).exists():
                        campotitulo = CamposTitulosPostulacion.objects.filter(status=True,
                                                                              titulo=f.cleaned_data['titulo']).first()
                    else:
                        campotitulo = CamposTitulosPostulacion(titulo=f.cleaned_data['titulo'])
                        campotitulo.save(request)
                    for ca in f.cleaned_data['campoamplio']:
                        if not campotitulo.campoamplio.filter(id=ca.id):
                            campotitulo.campoamplio.add(ca)
                    for ce in f.cleaned_data['campoespecifico']:
                        if not campotitulo.campoespecifico.filter(id=ce.id):
                            campotitulo.campoespecifico.add(ce)
                    for cd in f.cleaned_data['campodetallado']:
                        if not campotitulo.campodetallado.filter(id=cd.id):
                            campotitulo.campodetallado.add(cd)
                    campotitulo.save()
                    log(u'Adiciono titulacion: %s' % persona, request, "add")
                    return JsonResponse({'result': 'ok'})
                else:
                    return JsonResponse({'result': 'bad', 'mensaje': u'Error al guardar los datos'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', 'mensaje': u'Error al guardar los datos'})

        elif action == 'edittitulacionaspirante':
            try:
                persona = request.session['persona']
                f = TitulacionPersonaAdmisionForm(request.POST, request.FILES)
                if 'archivo' in request.FILES:
                    arch = request.FILES['archivo']
                    extencion = arch._name.split('.')
                    exte = extencion[1]
                    if arch.size > 4194304:
                        return JsonResponse({"result": "bad", "mensaje": u"Error, archivo mayor a 4 Mb."})
                    if not exte == 'pdf' and not exte == 'png' and not exte == 'jpg' and not exte == 'jpeg' and not exte == 'jpg':
                        return JsonResponse({"result": "bad", "mensaje": u"Solo archivos .pdf, jpg, jpeg, png"})

                if 'registroarchivo' in request.FILES:
                    registroarchivo = request.FILES['registroarchivo']
                    extencion1 = registroarchivo._name.split('.')
                    exte1 = extencion1[1]
                    if registroarchivo.size > 4194304:
                        return JsonResponse({"result": "bad", "mensaje": u"Error, archivo mayor a 4 Mb."})
                    if not exte1 == 'pdf' and not exte1 == 'png' and not exte1 == 'jpg' and not exte1 == 'jpeg' and not exte1 == 'jpg':
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"Archivo SENESCYT solo en formato .pdf, jpg, jpeg, png"})
                if f.is_valid():
                    titulacion = Titulacion.objects.get(pk=int(request.POST['id']))
                    titulacion.areatitulo = f.cleaned_data['areatitulo']
                    titulacion.fechainicio = f.cleaned_data['fechainicio']
                    titulacion.fechaobtencion = f.cleaned_data['fechaobtencion']
                    titulacion.fechaegresado = f.cleaned_data['fechaegresado']
                    # titulacion.areaconocimiento = f.cleaned_data['areaconocimiento']
                    # titulacion.subareaconocimiento = f.cleaned_data['subareaconocimiento']
                    # titulacion.subareaespecificaconocimiento = f.cleaned_data['subareaespecificaconocimiento']
                    titulacion.pais = f.cleaned_data['pais']
                    titulacion.provincia = f.cleaned_data['provincia']
                    titulacion.canton = f.cleaned_data['canton']
                    titulacion.parroquia = f.cleaned_data['parroquia']
                    titulacion.colegio = f.cleaned_data['colegio']
                    titulacion.anios = f.cleaned_data['anios']
                    titulacion.semestres = f.cleaned_data['semestres']
                    titulacion.aplicobeca = f.cleaned_data['aplicobeca']
                    titulacion.tipobeca = f.cleaned_data['tipobeca'] if f.cleaned_data['aplicobeca'] else None
                    titulacion.financiamientobeca = f.cleaned_data['financiamientobeca'] if f.cleaned_data[
                        'aplicobeca'] else None
                    titulacion.valorbeca = f.cleaned_data['valorbeca'] if f.cleaned_data['aplicobeca'] else 0
                    if not titulacion.verificadosenescyt:
                        titulacion.titulo = f.cleaned_data['titulo']
                        titulacion.educacionsuperior = f.cleaned_data['educacionsuperior']
                        titulacion.institucion = f.cleaned_data['institucion']
                        titulacion.registro = f.cleaned_data['registro']
                        titulacion.cursando = f.cleaned_data['cursando']
                    if 'archivo' in request.FILES:
                        newfile = request.FILES['archivo']
                        newfile._name = generar_nombre("titulacion_", newfile._name)
                        titulacion.archivo = newfile
                    if 'registroarchivo' in request.FILES:
                        newfile2 = request.FILES['registroarchivo']
                        if newfile2:
                            newfile2._name = generar_nombre("archivosenecyt_", newfile2._name)
                            titulacion.registroarchivo = newfile2
                    titulacion.save(request)
                    campotitulo = None
                    if CamposTitulosPostulacion.objects.filter(status=True, titulo=f.cleaned_data['titulo']).exists():
                        campotitulo = CamposTitulosPostulacion.objects.filter(status=True,
                                                                              titulo=f.cleaned_data['titulo']).first()
                    else:
                        campotitulo = CamposTitulosPostulacion(titulo=f.cleaned_data['titulo'])
                        campotitulo.save(request)
                    for ca in f.cleaned_data['campoamplio']:
                        if not campotitulo.campoamplio.filter(id=ca.id):
                            campotitulo.campoamplio.add(ca)
                    for ce in f.cleaned_data['campoespecifico']:
                        if not campotitulo.campoespecifico.filter(id=ce.id):
                            campotitulo.campoespecifico.add(ce)
                    for cd in f.cleaned_data['campodetallado']:
                        if not campotitulo.campodetallado.filter(id=cd.id):
                            campotitulo.campodetallado.add(cd)
                    campotitulo.save()
                    request.session['instruccion'] = 1
                    if Graduado.objects.filter(status=True, inscripcion__persona__id=persona.id).exists():
                        datos = Persona.objects.get(status=True, id=persona.id)
                        if 'personales' in request.session and 'nacimiento' in request.session and 'domicilio' in request.session and 'etnia' in request.session and 'instruccion' in request.session and datos.datosactualizados == 0:
                            if request.session['personales'] == 1 and request.session['nacimiento'] == 1 and \
                                    request.session['domicilio'] == 1 and request.session['etnia'] == 1 and \
                                    request.session['instruccion'] == 1 and datos.datosactualizados == 0:
                                datos.datosactualizados = 1
                                datos.save(request)
                    log(u'Modifico titulacion: %s' % persona, request, "edit")
                    return JsonResponse({'result': 'ok'})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', 'mensaje': u'Error al guardar los datos'})

        elif action == 'deltitulacionaspirante':
            try:
                persona = request.session['persona']
                titulacion = Titulacion.objects.get(pk=int(request.POST['id']))
                if titulacion.verificado:
                    return JsonResponse({'result': 'bad', 'mensaje': u'No puede eliminar el titulo.'})
                log(u'Elimino titulacion: %s' % titulacion, request, "del")
                titulacion.status = False
                titulacion.save()
                return JsonResponse({'result': True, 'error': False})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', 'mensaje': u'Error al eliminar los datos'})

        elif action == 'historialarchivopersonal':
            try:
                historial = None
                if HistorialArchivosContratos.objects.filter(personacontrato=int(request.POST['id']),
                                                             status=True).exists():
                    historial = HistorialArchivosContratos.objects.filter(personacontrato=int(request.POST['id']),
                                                                          status=True)
                data['historialarchivo'] = historial
                template = get_template("th_hojavida/historialcontratopersona.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'html': json_content})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'detallearticulo':
            try:
                data['articulo'] = articulos = ArticuloInvestigacion.objects.get(pk=request.POST['id'])
                data['evidencias'] = Evidencia.objects.filter(status=True, matrizevidencia_id=3)
                data['formevidencias'] = EvidenciaForm()
                # data['evidenciasarticulo'] = evidencias = DetalleEvidencias.objects.filter(articulo=articulos)
                template = get_template("th_hojavida/detallearticulo.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'html': json_content})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'detalleponencia':
            try:
                data['ponencias'] = ponencias = PonenciasInvestigacion.objects.get(pk=request.POST['id'])

                data['evidencias'] = Evidencia.objects.filter(status=True, matrizevidencia_id=4)
                data['formevidencias'] = EvidenciaForm()
                template = get_template("th_hojavida/detalleponencia.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'html': json_content})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'detallecapitulo':
            try:
                data['capitulos'] = capitulos = CapituloLibroInvestigacion.objects.get(pk=request.POST['id'])

                data['evidencias'] = Evidencia.objects.filter(status=True, matrizevidencia_id=7)
                data['formevidencias'] = EvidenciaForm()
                template = get_template("th_hojavida/detallecapitulo.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'html': json_content})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'detallelibro':
            try:
                data['libros'] = libros = LibroInvestigacion.objects.get(pk=request.POST['id'])

                data['evidencias'] = Evidencia.objects.filter(status=True, matrizevidencia_id=6)
                data['formevidencias'] = EvidenciaForm()
                template = get_template("th_hojavida/detalledetallelibro.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'html': json_content})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'addarchivotitulacion':
            try:
                persona = request.session['persona']
                f = ArchivoTitulacionForm(request.POST, request.FILES)
                if f.is_valid():
                    titulacion = Titulacion.objects.get(pk=int(request.POST['id']))
                    newfile = request.FILES['archivo']
                    newfile._name = generar_nombre("titulacion_", newfile._name)
                    titulacion.archivo = newfile
                    titulacion.save(request)
                    log(u'Adiciono archivo de titulacion: %s' % persona, request, "add")
                    return JsonResponse({'result': 'ok'})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', 'mensaje': u'Error al guardar los datos'})

        elif action == 'addarchivoidioma':
            try:
                persona = request.session['persona']
                f = ArchivoIdiomaForm(request.POST, request.FILES)
                if f.is_valid():
                    idioma = IdiomaDomina.objects.get(pk=int(request.POST['id']))
                    newfile = request.FILES['archivo']
                    newfile._name = generar_nombre("titulacion_", newfile._name)
                    idioma.archivo = newfile
                    idioma.save(request)
                    log(u'Adiciono archivo de idioma: %s' % persona, request, "add")
                    return JsonResponse({'result': 'ok'})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', 'mensaje': u'Error al guardar los datos'})

        elif action == 'addarchivocapacitacion':
            try:
                persona = request.session['persona']
                f = ArchivoTitulacionForm(request.POST, request.FILES)
                if f.is_valid():
                    titulacion = Capacitacion.objects.get(pk=int(request.POST['id']))
                    newfile = request.FILES['archivo']
                    newfile._name = generar_nombre("capacitacion_", newfile._name)
                    titulacion.archivo = newfile
                    titulacion.save(request)
                    log(u'Adiciono archivo de capacitacion: %s' % persona, request, "add")
                    return JsonResponse({'result': 'ok'})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', 'mensaje': u'Error al guardar los datos'})

        elif action == 'addgasto':
            try:
                persona = request.session['persona']
                datos = json.loads(request.POST['lista_items1'])
                if not PeriodoGastosPersonales.objects.filter(fechadesde__lte=datetime.now().date(),
                                                              fechahasta__gte=datetime.now().date()).exists():
                    return JsonResponse({'result': 'bad', 'mensaje': u'No existe un Periodo de actualización'})
                periodo = PeriodoGastosPersonales.objects.filter(fechadesde__lte=datetime.now().date(),
                                                                 fechahasta__gte=datetime.now().date())[0]

                gasto = persona.datos_actualizacion_gastos(periodo)

                gasto.rmuproyectado = Decimal(datos['rmu_proyectado'])
                gasto.rmupagado = Decimal(datos['rmu_pagado'])
                gasto.horasextraspagado = Decimal(datos['horas_extras_pagado'])
                gasto.horasextrasactual = Decimal(datos['horas_extras_actual'])
                gasto.horasextrasproyectado = Decimal(datos['horas_extras_proyectada'])
                gasto.otrosingresos = Decimal(datos['total_ingresos_con_otro'])
                gasto.totalingresos = Decimal(datos['total_anual_base'])
                gasto.otrosgastos = Decimal(datos['rmuproyectado'])
                gasto.rebajasotros = Decimal(datos['rmuproyectado'])
                gasto.totalgastos = Decimal(datos['total_gastos'])
                gasto.detallevivienda = Decimal(datos['vivienda'])
                gasto.detalleeducacion = Decimal(datos['educacion'])
                gasto.detallesalud = Decimal(datos['salud'])
                gasto.detallealimentacion = Decimal(datos['alimnentacion'])
                gasto.detallevestimenta = Decimal(datos['vestimenta'])
                gasto.excepcionesgastos = int(datos['excepcion'])

                # gasto.excepcionesgastos = Decimal(datos['rmuproyectado'])
                gasto.save(request)
                gasto.actualiza_detalle()
                log(u'Adiciono archivo de gastos: %s' % persona, request, "add")
                return JsonResponse({'result': 'ok'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', 'mensaje': u'Error al guardar los datos'})

        elif action == 'addarchivoexperiencia':
            try:
                persona = request.session['persona']
                f = ArchivoExperienciaForm(request.POST, request.FILES)
                if f.is_valid():
                    titulacion = ExperienciaLaboral.objects.get(pk=int(request.POST['id']))
                    newfile = request.FILES['archivo']
                    newfile._name = generar_nombre("experiencialaboral_", newfile._name)
                    titulacion.archivo = newfile
                    titulacion.save(request)
                    log(u'Adiciono archivo de referncia laboral: %s' % persona, request, "add")
                    return JsonResponse({'result': 'ok'})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', 'mensaje': u'Error al guardar los datos'})

        elif action == 'addarchivocontratofirmado':
            try:
                with transaction.atomic():
                    persona = request.session['persona']
                    contrato = PersonaContratos.objects.get(pk=int(request.POST['id']))
                    estado = int(request.POST['estado'])
                    if estado == 2 or estado == 5:
                        if not 'archivo' in request.FILES:
                            return JsonResponse({"result": True, "mensaje": "Subir contrato firmado en formato .pdf"},
                                                safe=False)

                    if 'archivo' in request.FILES:
                        d = request.FILES['archivo']
                        newfilesd = d._name
                        ext = newfilesd[newfilesd.rfind("."):]
                        if ext == '.pdf':
                            a = 1
                        else:
                            return JsonResponse({'result': True, "message": 'Error, solo archivos .pdf.'}, safe=False)
                        if d.size > 12582912:
                            return JsonResponse({'result': True, "message": 'Error, archivo mayor a 12 Mb.'},
                                                safe=False)

                    f = ArchivoHistorialContratoForm(request.POST, request.FILES)
                    if f.is_valid():
                        historial = HistorialArchivosContratos(personacontrato=contrato,
                                                               estado_archivo=2
                                                               )
                        historial.save(request)

                        newfile = None
                        if 'archivo' in request.FILES:
                            newfile = request.FILES['archivo']
                            newfile._name = generar_nombre("contratopersonalfirmado", newfile._name)
                            historial.archivo = newfile
                            historial.save(request)

                        contrato.ultimo_archivo = historial.id
                        contrato.subio_archivo = True
                        contrato.save(request)
                        # distributivo = DistributivoPersona.objects.filter(persona=contrato.persona,status=True).order_by("id")
                        # for dis in distributivo:
                        #     responsable = dis.unidadorganica.responsable

                        departamento = Departamento.objects.filter(permisodepartamento=2, status=True).order_by("id")
                        for depa in departamento:
                            if depa.responsable:
                                responsable = depa.responsable

                        if estado == 5:
                            notificacion = Notificacion(
                                titulo=f"Se ingreso contrato corregido y firmado por el personal {contrato.persona.nombre_completo()} Nro. Documento {contrato.numerodocumento}",
                                cuerpo=f"Se ingreso contrato corregido y firmado por el personal {contrato.persona.nombre_completo()} Nro. Documento {contrato.numerodocumento}, para su revisión y posteriormente ingresar el documento firmado.",
                                destinatario=responsable,
                                url=f"/th_personal?action=listadocontratosfirmados&s={contrato.persona.cedula}",
                                fecha_hora_visible=datetime.now() + timedelta(days=1),
                                content_type=ContentType.objects.get_for_model(contrato),
                                object_id=contrato.id,
                                prioridad=1,
                                app_label='sga'
                                )
                        else:
                            notificacion = Notificacion(
                                titulo=f"Se ingreso contrato firmado por el personal {contrato.persona.nombre_completo()} Nro. Documento {contrato.numerodocumento}",
                                cuerpo=f"Se ingreso contrato firmado por el personal {contrato.persona.nombre_completo()} Nro. Documento {contrato.numerodocumento}, para su revisión y posteriormente ingresar el documento firmado.",
                                destinatario=responsable,
                                url=f"/th_personal?action=listadocontratosfirmados&s={contrato.persona.cedula}",
                                fecha_hora_visible=datetime.now() + timedelta(days=1),
                                content_type=ContentType.objects.get_for_model(contrato),
                                object_id=contrato.id,
                                prioridad=1,
                                app_label='sga'
                                )
                        notificacion.save(request)
                        log(u'Adiciono contratro frimado: %s' % persona, request, "add")
                        return JsonResponse({"result": False}, safe=False)
                    else:
                        return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()],
                                             "message": "Error en el formulario"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        elif action == 'addidioma':
            try:
                persona = request.session['persona']
                form = IdiomaDominaForm(request.POST, request.FILES)
                if form.is_valid():
                    if persona.idiomadomina_set.filter(status=True, idioma=form.cleaned_data['idioma']).exists():
                        return JsonResponse({'result': 'bad', 'mensaje': u'Ya existe ese idioma registrado.'})
                    if form.cleaned_data['lenguamaterna']:
                        if persona.idiomadomina_set.filter(status=True, lenguamaterna=True).exists():
                            return JsonResponse(
                                {'result': 'bad', 'mensaje': u'Ya existe una lengua materna registrada.'})
                    idioma = IdiomaDomina(persona=persona,
                                          idioma=form.cleaned_data['idioma'],
                                          lectura=form.cleaned_data['lectura'],
                                          escritura=form.cleaned_data['escritura'],
                                          lenguamaterna=form.cleaned_data['lenguamaterna'],
                                          oral=form.cleaned_data['oral'])
                    idioma.save(request)
                    if 'archivo' in request.FILES:
                        newfile = request.FILES['archivo']
                        newfile._name = generar_nombre("idioma_", newfile._name)
                        idioma.archivo = newfile
                        idioma.save(request)
                    log(u'Adiciono idioma que domina: %s' % idioma, request, "add")
                    return JsonResponse({'result': 'ok'})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', 'mensaje': u'Error al guardar los datos'})

        elif action == 'editidioma':
            try:
                idioma = IdiomaDomina.objects.get(pk=request.POST['id'])
                form = IdiomaDominaForm(request.POST, request.FILES)
                if form.is_valid():
                    idioma.idioma = form.cleaned_data['idioma']
                    idioma.lectura = form.cleaned_data['lectura']
                    idioma.escritura = form.cleaned_data['escritura']
                    idioma.lenguamaterna = form.cleaned_data['lenguamaterna']
                    idioma.oral = form.cleaned_data['oral']
                    idioma.save(request)
                    if 'archivo' in request.FILES:
                        newfile = request.FILES['archivo']
                        newfile._name = generar_nombre("idioma_", newfile._name)
                        idioma.archivo = newfile
                        idioma.save(request)
                    return JsonResponse({'result': 'ok'})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', 'mensaje': u'Error al guardar los datos datos'})

        elif action == 'delidioma':
            try:
                idioma = IdiomaDomina.objects.get(pk=request.POST['id'])
                log(u"Elimino idioma que domina: %s" % idioma, request, "del")
                idioma.delete()
                return JsonResponse({'result': 'ok'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', 'mensaje': u'Error al eliminar los datos'})

        elif action == 'aplicaproyeccion':
            try:
                familiar = PersonaDatosFamiliares.objects.get(pk=encrypt_id(request.POST['id']))
                # if familiar.personaproyeccion and familiar.personaproyeccion == persona:
                #     raise NameError(f"No puede seleccionar esta carga familiar,la persona {familiar.personaproyeccion} la registró anteriormente.")
                if eval(request.POST['val'].capitalize()):
                    fam = PersonaDatosFamiliares.objects.filter(status=True, identificacion=familiar.identificacion, aplicaproyeccion=True).first()
                    if fam:
                        raise NameError(f"La carga familiar a sido seleccionado por: {fam.persona}")
                familiar.aplicaproyeccion = eval(request.POST['val'].capitalize())
                familiar.personaproyeccion = persona
                familiar.save(request)
                log(u"Seleccionó carga familiar para proyección de gastos: %s" % familiar, request, "del")
                return JsonResponse({'result': True, 'mensaje':'Guardado con exito'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': False, 'mensaje': f'{ex}'})

        elif action == 'detallerol':
            try:
                data['detallerol'] = registro = RolPago.objects.get(pk=int(request.POST['id']), status=True)
                data['detalleinformativo'] = registro.detallerolinformativo()
                data['detalleingreso'] = registro.detallerolingreso()
                data['detalleegreso'] = registro.detallerolegreso_consolidado()
                template = get_template("th_nomina/detallerol.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'html': json_content})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'agregarinstitucion':
            try:
                f = InstitucionForm(request.POST)
                if f.is_valid():
                    nombre = f.cleaned_data['nombreinstitucion']
                    if InstitucionBeca.objects.filter(nombre=nombre, status=True, tiporegistro=1).exists():
                        return JsonResponse({"result": "bad", "mensaje": u"El nombre de la institución ya existe."})
                    else:
                        institucion = InstitucionBeca(nombre=nombre, tiporegistro=1)
                        institucion.save(request)
                        log(u'Agrego institución: %s' % (institucion), request, "add")
                        return JsonResponse({"result": "ok", "id": institucion.id})
                else:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al enviar el formulario."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'addpublicacion':
            try:
                f = SolicitudPublicacionForm(request.POST, request.FILES)

                if 'archivo' in request.FILES:
                    newfile = request.FILES['archivo']
                    extension = newfile._name.split('.')
                    tam = len(extension)
                    exte = extension[tam - 1]
                    if newfile.size > 10485760:
                        return JsonResponse({"result": "bad",
                                             "mensaje": u"Error, Tamaño de archivo Publicación Máximo permitido es de 10Mb"})
                    if exte.lower() not in ['pdf', 'doc', 'docx']:
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"Solo se permiten archivos .pdf, .doc y .docx (Publicación)"})

                if 'archivocertificado' in request.FILES:
                    newfile = request.FILES['archivocertificado']
                    extension = newfile._name.split('.')
                    tam = len(extension)
                    exte = extension[tam - 1]
                    if newfile.size > 10485760:
                        return JsonResponse({"result": "bad",
                                             "mensaje": u"Error, Tamaño de archivo Carta de Aceptación Máximo permitido es de 10Mb"})
                    if exte.lower() not in ['pdf', 'doc', 'docx']:
                        return JsonResponse({"result": "bad",
                                             "mensaje": u"Solo se permiten archivos .pdf, .doc y .docx (Carta de Aceptación)"})

                if 'archivoparticipacion' in request.FILES:
                    newfile = request.FILES['archivoparticipacion']
                    extension = newfile._name.split('.')
                    tam = len(extension)
                    exte = extension[tam - 1]
                    if newfile.size > 10485760:
                        return JsonResponse({"result": "bad",
                                             "mensaje": u"Error, Tamaño de archivo Certificado de Participación Máximo permitido es de 10Mb"})
                    if exte.lower() not in ['pdf', 'doc', 'docx']:
                        return JsonResponse({"result": "bad",
                                             "mensaje": u"Solo se permiten archivos .pdf, .doc y .docx (Certificado de Participación)"})

                if 'archivocomite' in request.FILES:
                    newfile = request.FILES['archivocomite']
                    extension = newfile._name.split('.')
                    tam = len(extension)
                    exte = extension[tam - 1]
                    if newfile.size > 10485760:
                        return JsonResponse({"result": "bad",
                                             "mensaje": u"Error, Tamaño de archivo Comité Científico Evaluador Máximo permitido es de 10Mb"})
                    if exte.lower() not in ['pdf', 'doc', 'docx']:
                        return JsonResponse({"result": "bad",
                                             "mensaje": u"Solo se permiten archivos .pdf, .doc y .docx (Comité Científico Evaluador)"})

                if f.is_valid():
                    if int(f.cleaned_data['tiposolicitud']) == 1 or int(f.cleaned_data['tiposolicitud']) == 5:
                        if int(f.cleaned_data['estadopublicacion']) == 1:
                            publicacion = request.FILES['archivo']
                            cartaaceptacion = request.FILES['archivocertificado']

                            if publicacion._name == cartaaceptacion._name:
                                return JsonResponse({"result": "bad",
                                                     "mensaje": u"Los archivos de publicación y carta de aceptación no deben ser iguales"})

                        if SolicitudPublicacion.objects.filter(status=True,
                                                               tiposolicitud=f.cleaned_data['tiposolicitud'],
                                                               nombre=f.cleaned_data['nombre']).exists():
                            return JsonResponse(
                                {"result": "bad", "mensaje": u"El título para ese tipo de publicación ya existe"})

                        if SolicitudPublicacion.objects.filter(status=True,
                                                               tiposolicitud=f.cleaned_data['tiposolicitud'],
                                                               motivo=f.cleaned_data['motivo']).exists():
                            return JsonResponse(
                                {"result": "bad", "mensaje": u"El resumen para ese tipo de publicación ya existe"})

                    estadosolicitud = obtener_estado_solicitud(8, 1)  # SOLICITADO

                    if int(f.cleaned_data['tiposolicitud']) == 1 or int(f.cleaned_data['tiposolicitud']) == 5:
                        solicitudpublicacion = SolicitudPublicacion(persona=persona,
                                                                    tiposolicitud=f.cleaned_data['tiposolicitud'],
                                                                    nombre=f.cleaned_data['nombre'],
                                                                    motivo=f.cleaned_data['motivo'],
                                                                    revistainvestigacion=f.cleaned_data['revista2'],
                                                                    estadopublicacion=f.cleaned_data[
                                                                        'estadopublicacion'],
                                                                    fecharecepcion=f.cleaned_data['fecharecepcion'],
                                                                    fechaaprobacion=f.cleaned_data['fechaaprobacion'],
                                                                    fechapublicacion=f.cleaned_data[
                                                                        'fechapublicacion'] if int(f.cleaned_data[
                                                                                                       'estadopublicacion']) == 1 else None,
                                                                    enlace=f.cleaned_data['enlace'] if int(
                                                                        f.cleaned_data[
                                                                            'estadopublicacion']) == 1 else '',
                                                                    volumen=f.cleaned_data['volumen'] if int(
                                                                        f.cleaned_data[
                                                                            'estadopublicacion']) == 1 else '',
                                                                    numero=f.cleaned_data['numero'] if int(
                                                                        f.cleaned_data[
                                                                            'estadopublicacion']) == 1 else '',
                                                                    paginas=f.cleaned_data['paginas'] if int(
                                                                        f.cleaned_data[
                                                                            'estadopublicacion']) == 1 else '',
                                                                    areaconocimiento=f.cleaned_data['areaconocimiento'],
                                                                    subareaconocimiento=f.cleaned_data[
                                                                        'subareaconocimiento'],
                                                                    subareaespecificaconocimiento=f.cleaned_data[
                                                                        'subareaespecificaconocimiento'],
                                                                    lineainvestigacion=f.cleaned_data[
                                                                        'lineainvestigacion'],
                                                                    sublineainvestigacion=f.cleaned_data[
                                                                        'sublineainvestigacion'],
                                                                    provieneproyecto=f.cleaned_data['provieneproyecto'],
                                                                    tipoproyecto=f.cleaned_data['tipoproyecto'] if
                                                                    f.cleaned_data['provieneproyecto'] else None,
                                                                    estado=estadosolicitud
                                                                    )
                    elif int(f.cleaned_data['tiposolicitud']) == 2:
                        if SolicitudPublicacion.objects.filter(status=True,
                                                               tiposolicitud=f.cleaned_data['tiposolicitud'],
                                                               nombre=f.cleaned_data['nombre']).exists():
                            return JsonResponse(
                                {"result": "bad", "mensaje": u"El título para ese tipo de publicación ya existe"})

                        if SolicitudPublicacion.objects.filter(status=True,
                                                               tiposolicitud=f.cleaned_data['tiposolicitud'],
                                                               motivo=f.cleaned_data['motivo']).exists():
                            return JsonResponse(
                                {"result": "bad", "mensaje": u"El resumen para ese tipo de publicación ya existe"})

                        if f.cleaned_data['fecharecepcion'] > f.cleaned_data['fechaaprobacion']:
                            return JsonResponse({"result": "bad",
                                                 "mensaje": u"La fecha de inicio deber ser menor o igual a la fecha fin"})

                        if int(f.cleaned_data['estadopublicacionponencia']) == 1:
                            if f.cleaned_data['fechapublicacionponencia'] >= f.cleaned_data['fechaaprobacion']:
                                return JsonResponse({"result": "bad",
                                                     "mensaje": u"La fecha de publicación debe ser menor que la fecha fin del congreso"})

                        integrantes = request.POST.getlist('integrante[]')
                        instituciones = request.POST.getlist('institucion[]')
                        emails = request.POST.getlist('email[]')

                        if f.cleaned_data['comitecientifico'] is True:
                            if not integrantes:
                                return JsonResponse({"result": "bad",
                                                     "mensaje": u"Ingrese los datos de los integrantes del comité científico evaluador"})

                            if len(integrantes) < 3:
                                return JsonResponse(
                                    {"result": "bad", "mensaje": u"Ingrese mínimo 3 integrantes del comité"})

                        if integrantes:
                            nenblanco = [nombre for nombre in integrantes if nombre.strip() == '']
                            ienblanco = [institucion for institucion in instituciones if institucion.strip() == '']
                            emailnovalido = [email for email in emails if
                                             email.strip() != '' and not email_valido(email)]
                            if nenblanco or ienblanco:
                                return JsonResponse({"result": "bad",
                                                     "mensaje": u"Los datos de los integrantes del comité científico evaluador son obligatorios a excepción del e-mail"})

                            if emailnovalido:
                                return JsonResponse({"result": "bad",
                                                     "mensaje": u"El formato de una o varias direcciones de e-mail de los integrantes del comité científico evaluador no es válido"})

                            listado = [nombre.strip() for nombre in integrantes if nombre.strip() != '']
                            repetido = [x for x, y in collections.Counter(listado).items() if y > 1]
                            if repetido:
                                return JsonResponse({"result": "bad",
                                                     "mensaje": u"El integrante [%s] del comité científico evaluador está repetido" % (
                                                         repetido[0])})

                        solicitudpublicacion = SolicitudPublicacion(persona=persona,
                                                                    tiposolicitud=f.cleaned_data['tiposolicitud'],
                                                                    nombre=f.cleaned_data['nombre'],
                                                                    motivo=f.cleaned_data['motivo'],
                                                                    evento=f.cleaned_data['evento'],
                                                                    pais=f.cleaned_data['pais'],
                                                                    ciudad=f.cleaned_data['ciudad'],
                                                                    fecharecepcion=f.cleaned_data['fecharecepcion'],
                                                                    fechaaprobacion=f.cleaned_data['fechaaprobacion'],
                                                                    fechapublicacion=f.cleaned_data[
                                                                        'fechapublicacionponencia'] if int(
                                                                        f.cleaned_data[
                                                                            'estadopublicacionponencia']) == 1 else None,
                                                                    estadopublicacion=f.cleaned_data[
                                                                        'estadopublicacionponencia'],
                                                                    enlace=f.cleaned_data['enlace'],
                                                                    areaconocimiento=f.cleaned_data['areaconocimiento'],
                                                                    subareaconocimiento=f.cleaned_data[
                                                                        'subareaconocimiento'],
                                                                    subareaespecificaconocimiento=f.cleaned_data[
                                                                        'subareaespecificaconocimiento'],
                                                                    lineainvestigacion=f.cleaned_data[
                                                                        'lineainvestigacion'],
                                                                    sublineainvestigacion=f.cleaned_data[
                                                                        'sublineainvestigacion'],
                                                                    provieneproyecto=f.cleaned_data['provieneproyecto'],
                                                                    tipoproyecto=f.cleaned_data['tipoproyecto'] if
                                                                    f.cleaned_data['provieneproyecto'] else None,
                                                                    comitecientifico=f.cleaned_data['comitecientifico'],
                                                                    estado=estadosolicitud
                                                                    )
                    elif int(f.cleaned_data['tiposolicitud']) == 3:
                        solicitudpublicacion = SolicitudPublicacion(persona=persona,
                                                                    tiposolicitud=f.cleaned_data['tiposolicitud'],
                                                                    motivo=f.cleaned_data['motivo'],
                                                                    nombre=f.cleaned_data['nombre'],
                                                                    pais=f.cleaned_data['pais'],
                                                                    ciudad=f.cleaned_data['ciudad'],
                                                                    fechapublicacion=f.cleaned_data['fechapublicacion'],
                                                                    areaconocimiento=f.cleaned_data['areaconocimiento'],
                                                                    subareaconocimiento=f.cleaned_data[
                                                                        'subareaconocimiento'],
                                                                    subareaespecificaconocimiento=f.cleaned_data[
                                                                        'subareaespecificaconocimiento'],
                                                                    estado=estadosolicitud)
                    else:
                        solicitudpublicacion = SolicitudPublicacion(persona=persona,
                                                                    tiposolicitud=f.cleaned_data['tiposolicitud'],
                                                                    motivo=f.cleaned_data['motivo'],
                                                                    estado=estadosolicitud)

                    solicitudpublicacion.save(request)

                    if int(f.cleaned_data['tiposolicitud']) in [1, 5, 2]:
                        if f.cleaned_data['provieneproyecto']:
                            if int(f.cleaned_data['tipoproyecto']) != 3:
                                solicitudpublicacion.proyectointerno = f.cleaned_data['proyectointerno']
                            else:
                                solicitudpublicacion.proyectoexterno = f.cleaned_data['proyectoexterno']

                    if 'archivo' in request.FILES:
                        newfile = request.FILES['archivo']
                        newfile._name = generar_nombre("solicitudpublicacion_", newfile._name)
                        solicitudpublicacion.archivo = newfile

                    if 'archivocertificado' in request.FILES:
                        newfile = request.FILES['archivocertificado']
                        newfile._name = generar_nombre("solicitudpublicacioncertificacion_", newfile._name)
                        solicitudpublicacion.archivocertificado = newfile

                    if 'archivoparticipacion' in request.FILES:
                        newfile = request.FILES['archivoparticipacion']
                        newfile._name = generar_nombre("solicitudpublicacionparticipacion_", newfile._name)
                        solicitudpublicacion.archivoparticipacion = newfile

                    if 'archivocomite' in request.FILES:
                        newfile = request.FILES['archivocomite']
                        newfile._name = generar_nombre("solicitudpublicacioncomite_", newfile._name)
                        solicitudpublicacion.archivocomite = newfile

                    if int(f.cleaned_data['tiposolicitud']) == 2:
                        integrantes = request.POST.getlist('integrante[]')
                        instituciones = request.POST.getlist('institucion[]')
                        emails = request.POST.getlist('email[]')

                        comiteevaluador = "|".join([
                            integrante.upper().strip() + "," + institucion.upper().strip() + "," + email.lower().strip()
                            for integrante, institucion, email in
                            zip(integrantes, instituciones, emails)])
                        solicitudpublicacion.integrantecomite = comiteevaluador

                    solicitudpublicacion.save(request)

                    lista = ['investigacion@unemi.edu.ec']
                    send_html_mail("Solicitud Publicación",
                                   "emails/solicitudpublicacion.html",
                                   {'sistema': request.session['nombresistema'],
                                    'persona': persona,
                                    'fecha': datetime.now().date(),
                                    'hora': datetime.now().time(),
                                    'tipopublicacion': solicitudpublicacion.get_tiposolicitud_display(),
                                    't': tituloinstitucion()},
                                   lista,
                                   [],
                                   cuenta=CUENTAS_CORREOS[1][1])
                    log(u'Ingreso Solicitud Publicación: %s' % persona, request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                    errorformulario = f._errors
                    raise NameError('Error')
            except Exception as ex:
                msg = ex.__str__()
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos. Detalle: %s" % (msg)})

        elif action == 'addcapitulolibro':
            try:
                f = CapituloLibroHojaVidaForm(request.POST, request.FILES)

                archivolibro = request.FILES['archivolibro']
                descripcionarchivo = 'Libro'
                resp = validar_archivo(descripcionarchivo, archivolibro, ['pdf', 'doc', 'docx'], '10MB')
                if resp['estado'] != "OK":
                    return JsonResponse(
                        {"result": "bad", "titulo": "Atención!!!", "mensaje": resp["mensaje"], "showSwal": "True",
                         "swalType": "warning"})

                archivocapitulo = request.FILES['archivocapitulo']
                descripcionarchivo = 'Capítulo de libro'
                resp = validar_archivo(descripcionarchivo, archivocapitulo, ['pdf', 'doc', 'docx'], '10MB')
                if resp['estado'] != "OK":
                    return JsonResponse(
                        {"result": "bad", "titulo": "Atención!!!", "mensaje": resp["mensaje"], "showSwal": "True",
                         "swalType": "warning"})

                archivocarta = request.FILES['archivocarta']
                descripcionarchivo = 'Carta de aceptación'
                resp = validar_archivo(descripcionarchivo, archivocarta, ['pdf', 'doc', 'docx'], '10MB')
                if resp['estado'] != "OK":
                    return JsonResponse(
                        {"result": "bad", "titulo": "Atención!!!", "mensaje": resp["mensaje"], "showSwal": "True",
                         "swalType": "warning"})

                archivopares = request.FILES['archivopares']
                descripcionarchivo = 'Informe revsión pares'
                resp = validar_archivo(descripcionarchivo, archivopares, ['pdf', 'doc', 'docx'], '10MB')
                if resp['estado'] != "OK":
                    return JsonResponse(
                        {"result": "bad", "titulo": "Atención!!!", "mensaje": resp["mensaje"], "showSwal": "True",
                         "swalType": "warning"})

                if f.is_valid():
                    # Verifico que no exista el título del capítulo
                    if SolicitudPublicacion.objects.filter(status=True, tiposolicitud=4,
                                                           nombre=f.cleaned_data['titulocapitulo']).exists():
                        return JsonResponse({"result": "bad", "titulo": "Atención!!!",
                                             "mensaje": "El capítulo de libro ya ha sido ingresado", "showSwal": "True",
                                             "swalType": "warning"})

                    estadosolicitud = obtener_estado_solicitud(8, 1)  # SOLICITADO

                    archivolibrog = archivolibro
                    archivolibrog._name = generar_nombre("libro", archivolibrog._name)

                    archivocapitulog = archivocapitulo
                    archivocapitulog._name = generar_nombre("capitulolibro", archivocapitulog._name)

                    archivocartag = archivocarta
                    archivocartag._name = generar_nombre("cartaaceptacion", archivocartag._name)

                    archivoparesg = archivopares
                    archivoparesg._name = generar_nombre("revisionpares", archivoparesg._name)

                    solicitudpublicacion = SolicitudPublicacion(
                        persona=persona,
                        tiposolicitud=4,
                        nombre=f.cleaned_data['titulocapitulo'],
                        motivo=f.cleaned_data['resumen'],
                        evento=f.cleaned_data['titulolibro'],
                        codigoisbn=f.cleaned_data['codigoisbn'],
                        paginas=f.cleaned_data['paginas'],
                        editorcompilador=f.cleaned_data['editorcompilador'],
                        fechapublicacion=f.cleaned_data['fechapublicacion'],
                        filiacion=f.cleaned_data['filiacion'],
                        areaconocimiento=f.cleaned_data['areaconocimiento'],
                        subareaconocimiento=f.cleaned_data['subareaconocimiento'],
                        subareaespecificaconocimiento=f.cleaned_data['subareaespecificaconocimiento'],
                        archivo=archivocapitulog,
                        archivocertificado=archivolibrog,
                        archivoparticipacion=archivocartag,
                        archivocomite=archivoparesg,
                        estado=estadosolicitud
                    )
                    solicitudpublicacion.save(request)

                    lista = ['investigacion.dip@unemi.edu.ec']

                    send_html_mail("Solicitud Publicación",
                                   "emails/solicitudpublicacion.html",
                                   {'sistema': request.session['nombresistema'],
                                    'persona': persona,
                                    'fecha': datetime.now().date(),
                                    'hora': datetime.now().time(),
                                    'tipopublicacion': solicitudpublicacion.get_tiposolicitud_display(),
                                    't': tituloinstitucion()},
                                   lista,
                                   [],
                                   cuenta=CUENTAS_CORREOS[1][1])

                    log(u'% s ingresó solicitud de publicación: %s' % (persona, solicitudpublicacion), request, "add")
                    return JsonResponse(
                        {"result": "ok", "titulo": "Proceso exitoso!!!", "mensaje": u"Registro guardado con éxito",
                         "showSwal": True})
                else:
                    errorformulario = f._errors
                    raise NameError('Error en el formulario')
            except Exception as ex:
                msg = ex.__str__()
                transaction.set_rollback(True)
                return JsonResponse(
                    {"result": "bad", "titulo": "Error", "mensaje": u"Error al guardar los datos. [%s]" % msg,
                     "showSwal": "True", "swalType": "error"})

        elif action == 'addredacademica':
            try:
                f = RedAcademicaForm(request.POST)

                if not f.is_valid():
                    for k, v in f.errors.items():
                        raise NameError( k + ', ' + v[0])

                redacademica = RedPersona(
                    persona=persona,
                    tipo=f.cleaned_data['tipo'],
                    enlace=f.cleaned_data['enlace']
                )
                redacademica.save(request)
                log(u'%s agregó red académica: %s' % (persona, redacademica), request, "add")

                return JsonResponse({"result": False}, safe=False)
            except Exception as ex:
                msg = ex.__str__()
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos. [%s]" % (msg)})

        elif action == 'editredacademica':
            try:
                f = RedAcademicaForm(request.POST)

                if not f.is_valid():
                    for k, v in f.errors.items():
                        raise NameError( k + ', ' + v[0])

                # Consulto red académica
                redacademica = RedPersona.objects.get(pk=int(encrypt(request.POST['id'])))

                # Actualizo red académica
                redacademica.tipo = f.cleaned_data['tipo']
                redacademica.enlace = f.cleaned_data['enlace']
                redacademica.save(request)

                log(u'%s editó red académica: %s' % (persona, redacademica), request, "add")

                return JsonResponse({"result": False}, safe=False)
            except Exception as ex:
                msg = ex.__str__()
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos. [%s]" % (msg)})

        elif action == 'delredacademica':
            try:
                # Consulto la red académica
                redacademica = RedPersona.objects.get(pk=int(encrypt(request.POST['id'])))

                # Elimino la red académica
                redacademica.status = False
                redacademica.save(request)

                log(u'%s eliminó red académica: %s' % (persona, redacademica), request, "del")
                res_js = {'error': False}
            except Exception as ex:
                transaction.set_rollback(True)
                err = 'Error on line {}'.format(sys.exc_info()[-1].tb_lineno)
                msg_err = f"Ocurrio un error: {ex.__str__()}. {err}"
                res_js = {'error': True, 'message': msg_err}
            return JsonResponse(res_js)

        elif action == 'addparrevisor':
            try:
                f = ParRevisorArticuloForm(request.POST, request.FILES)
                if not f.is_valid():
                    for k, v in f.errors.items():
                        raise NameError( k + ', ' + v[0])

                if not 'archivo' in request.FILES:
                    raise NameError('Debe elegir el archivo del certificado en formato PDF')

                newfile = request.FILES['archivo']
                newfile._name = generar_nombre("certrevarticulo", newfile._name)

                parrevisor = ParRevisorProduccionCientifica(
                    persona=persona,
                    tipoproduccion=1,
                    fecharevision=f.cleaned_data['fecharevision'],
                    titulo=f.cleaned_data['titulo'],
                    revista=f.cleaned_data['revista'],
                    archivo=newfile
                )
                parrevisor.save(request)

                migrar_par_revisor(parrevisor, persona, periodoacademico, request)
                log(u'%s agregó par revisor de artículos: %s' % (persona, parrevisor), request, "add")
                return JsonResponse({"result": False}, safe=False)
            except Exception as ex:
                msg = ex.__str__()
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos. [%s]" % (msg)})

        elif action == 'editparrevisor':
            try:
                f = ParRevisorArticuloForm(request.POST, request.FILES)

                if not f.is_valid():
                    for k, v in f.errors.items():
                        raise NameError( k + ', ' + v[0])

                # Consulto par revisor
                parrevisor = ParRevisorProduccionCientifica.objects.get(pk=int(encrypt(request.POST['id'])))

                # Actualizo par revisor
                parrevisor.fecharevision = f.cleaned_data['fecharevision']
                parrevisor.titulo = f.cleaned_data['titulo']
                parrevisor.revista = f.cleaned_data['revista']

                if 'archivo' in request.FILES:
                    newfile = request.FILES['archivo']
                    newfile._name = generar_nombre("certrevarticulo", newfile._name)
                    parrevisor.archivo = newfile

                parrevisor.save(request)

                log(u'%s editó par revisor de artículos: %s' % (persona, parrevisor), request, "edit")

                migrar_par_revisor(parrevisor, persona, periodoacademico, request)
                return JsonResponse({"result": False}, safe=False)
            except Exception as ex:
                msg = ex.__str__()
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos. [%s]" % (msg)})

        elif action == 'delparrevisor':
            try:
                # Consulto el registro de par revisor
                parrevisor = ParRevisorProduccionCientifica.objects.get(pk=int(encrypt(request.POST['id'])))

                # Elimino el registro de par revisor
                parrevisor.status = False
                parrevisor.save(request)

                # Elimino la evidencia asociada
                if migracion := parrevisor.migracionevidenciaactividad_set.first():
                    if migracion.evidencia.estadoaprobacion == 1:
                        migracion.evidencia.delete()
                        migracion.delete()

                log(u'%s eliminó par revisor de artículo: %s' % (persona, parrevisor), request, "del")
                res_js = {'error': False}
            except Exception as ex:
                transaction.set_rollback(True)
                err = 'Error on line {}'.format(sys.exc_info()[-1].tb_lineno)
                msg_err = f"Ocurrio un error: {ex.__str__()}. {err}"
                res_js = {'error': True, 'message': msg_err}
            return JsonResponse(res_js)

        elif action == 'delsolicitud':
            try:
                solicitud = SolicitudPublicacion.objects.get(pk=request.POST['id'])
                solicitud.status = False
                solicitud.save(request)
                log(u'Elimino solicitud: %s' % solicitud, request, "del")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        elif action == 'editsolicitud':
            try:
                f = SolicitudPublicacionForm(request.POST, request.FILES)
                if 'archivo' in request.FILES:
                    newfile = request.FILES['archivo']
                    extension = newfile._name.split('.')
                    tam = len(extension)
                    exte = extension[tam - 1]
                    if newfile.size > 10485760:
                        return JsonResponse({"result": "bad",
                                             "mensaje": u"Error, Tamaño de archivo Publicación Máximo permitido es de 10Mb"})
                    if exte.lower() not in ['pdf', 'doc', 'docx']:
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"Solo se permiten archivos .pdf, .doc y .docx (Publicación)"})

                if 'archivocertificado' in request.FILES:
                    newfile = request.FILES['archivocertificado']
                    extension = newfile._name.split('.')
                    tam = len(extension)
                    exte = extension[tam - 1]
                    if newfile.size > 10485760:
                        return JsonResponse({"result": "bad",
                                             "mensaje": u"Error, Tamaño de archivo Carta de Aceptación Máximo permitido es de 10Mb"})
                    if exte.lower() not in ['pdf', 'doc', 'docx']:
                        return JsonResponse({"result": "bad",
                                             "mensaje": u"Solo se permiten archivos .pdf, .doc y .docx (Carta de Aceptación)"})

                if 'archivoparticipacion' in request.FILES:
                    newfile = request.FILES['archivoparticipacion']
                    extension = newfile._name.split('.')
                    tam = len(extension)
                    exte = extension[tam - 1]
                    if newfile.size > 10485760:
                        return JsonResponse({"result": "bad",
                                             "mensaje": u"Error, Tamaño de archivo Certificado de Participación Máximo permitido es de 10Mb"})
                    if exte.lower() not in ['pdf', 'doc', 'docx']:
                        return JsonResponse({"result": "bad",
                                             "mensaje": u"Solo se permiten archivos .pdf, .doc y .docx (Certificado de Participación)"})

                if 'archivocomite' in request.FILES:
                    newfile = request.FILES['archivocomite']
                    extension = newfile._name.split('.')
                    tam = len(extension)
                    exte = extension[tam - 1]
                    if newfile.size > 10485760:
                        return JsonResponse({"result": "bad",
                                             "mensaje": u"Error, Tamaño de archivo Comité Científico Evaluador Máximo permitido es de 10Mb"})
                    if exte.lower() not in ['pdf', 'doc', 'docx']:
                        return JsonResponse({"result": "bad",
                                             "mensaje": u"Solo se permiten archivos .pdf, .doc y .docx (Comité Científico Evaluador)"})

                if f.is_valid():
                    solicitud = SolicitudPublicacion.objects.get(pk=request.POST['id'])
                    estadosolicitud = obtener_estado_solicitud(8, 1)  # SOLICITADO

                    if solicitud.tiposolicitud == 1 or solicitud.tiposolicitud == 5:
                        if int(f.cleaned_data['estadopublicacion']) == 1:
                            if 'archivo' in request.FILES and 'archivocertificado' in request.FILES:
                                publicacion = request.FILES['archivo']
                                cartaaceptacion = request.FILES['archivocertificado']

                                if publicacion._name == cartaaceptacion._name:
                                    return JsonResponse({"result": "bad",
                                                         "mensaje": u"Los archivos de publicación y carta de aceptación no deben ser iguales"})

                        if SolicitudPublicacion.objects.filter(status=True, tiposolicitud=solicitud.tiposolicitud,
                                                               nombre=f.cleaned_data['nombre']).exclude(
                                pk=request.POST['id']).exists():
                            return JsonResponse(
                                {"result": "bad", "mensaje": u"El título para ese tipo de publicación ya existe"})

                        if SolicitudPublicacion.objects.filter(status=True, tiposolicitud=solicitud.tiposolicitud,
                                                               motivo=f.cleaned_data['motivo']).exclude(
                                pk=request.POST['id']).exists():
                            return JsonResponse(
                                {"result": "bad", "mensaje": u"El detalle para ese tipo de publicación ya existe"})

                        # solicitud = Solicitu  dPublicacion.objects.get(pk=request.POST['id'])
                        # if solicitud.tiposolicitud == 1 or solicitud.tiposolicitud == 5:
                        solicitud.nombre = f.cleaned_data['nombre']
                        solicitud.motivo = f.cleaned_data['motivo']
                        solicitud.revistainvestigacion = f.cleaned_data['revista2']
                        solicitud.estadopublicacion = f.cleaned_data['estadopublicacion']
                        solicitud.fecharecepcion = f.cleaned_data['fecharecepcion']
                        solicitud.fechaaprobacion = f.cleaned_data['fechaaprobacion']
                        solicitud.observacion = ""
                        solicitud.estado = estadosolicitud

                        if int(f.cleaned_data['estadopublicacion']) == 1:
                            solicitud.fechapublicacion = f.cleaned_data['fechapublicacion']
                            solicitud.enlace = f.cleaned_data['enlace']
                            solicitud.volumen = f.cleaned_data['volumen']
                            solicitud.numero = f.cleaned_data['numero']
                            solicitud.paginas = f.cleaned_data['paginas']
                        else:
                            solicitud.fechapublicacion = None
                            solicitud.enlace = ''
                            solicitud.volumen = ''
                            solicitud.numero = ''
                            solicitud.paginas = ''

                        solicitud.areaconocimiento = f.cleaned_data['areaconocimiento']
                        solicitud.subareaconocimiento = f.cleaned_data['subareaconocimiento']
                        solicitud.subareaespecificaconocimiento = f.cleaned_data['subareaespecificaconocimiento']
                        solicitud.lineainvestigacion = f.cleaned_data['lineainvestigacion']
                        solicitud.sublineainvestigacion = f.cleaned_data['sublineainvestigacion']
                        solicitud.provieneproyecto = f.cleaned_data['provieneproyecto']
                        solicitud.tipoproyecto = f.cleaned_data['tipoproyecto'] if f.cleaned_data[
                            'provieneproyecto'] else None

                        if f.cleaned_data['provieneproyecto']:
                            if int(f.cleaned_data['tipoproyecto']) != 3:
                                solicitud.proyectointerno = f.cleaned_data['proyectointerno']
                                solicitud.proyectoexterno = None
                            else:
                                solicitud.proyectoexterno = f.cleaned_data['proyectoexterno']
                                solicitud.proyectointerno = None
                        else:
                            solicitud.proyectoexterno = None
                            solicitud.proyectointerno = None

                    elif solicitud.tiposolicitud == 2:
                        if SolicitudPublicacion.objects.filter(status=True, tiposolicitud=solicitud.tiposolicitud,
                                                               nombre=f.cleaned_data['nombre']).exclude(
                                pk=request.POST['id']).exists():
                            return JsonResponse(
                                {"result": "bad", "mensaje": u"El título para ese tipo de publicación ya existe"})

                        if SolicitudPublicacion.objects.filter(status=True, tiposolicitud=solicitud.tiposolicitud,
                                                               motivo=f.cleaned_data['motivo']).exclude(
                                pk=request.POST['id']).exists():
                            return JsonResponse(
                                {"result": "bad", "mensaje": u"El resumen para ese tipo de publicación ya existe"})

                        if f.cleaned_data['fecharecepcion'] > f.cleaned_data['fechaaprobacion']:
                            return JsonResponse({"result": "bad",
                                                 "mensaje": u"La fecha de inicio deber ser menor o igual a la fecha fin"})

                        if int(f.cleaned_data['estadopublicacionponencia']) == 1:
                            if f.cleaned_data['fechapublicacionponencia'] >= f.cleaned_data['fechaaprobacion']:
                                return JsonResponse({"result": "bad",
                                                     "mensaje": u"La fecha de publicación debe ser menor que la fecha fin del congreso"})

                        integrantes = request.POST.getlist('integrante[]')
                        instituciones = request.POST.getlist('institucion[]')
                        emails = request.POST.getlist('email[]')

                        if f.cleaned_data['comitecientifico'] is True:
                            if not integrantes:
                                return JsonResponse({"result": "bad",
                                                     "mensaje": u"Ingrese los datos de los integrantes del comité científico evaluador"})

                            if len(integrantes) < 3:
                                return JsonResponse(
                                    {"result": "bad", "mensaje": u"Ingrese mínimo 3 integrantes del comité"})

                        # if f.cleaned_data['comitecientifico'] is True and not integrantes:
                        #     return JsonResponse({"result": "bad", "mensaje": u"Ingrese los datos de los integrantes del comité científico evaluador"})

                        if f.cleaned_data['comitecientifico'] is True and integrantes:
                            nenblanco = [nombre for nombre in integrantes if nombre.strip() == '']
                            ienblanco = [institucion for institucion in instituciones if institucion.strip() == '']
                            emailnovalido = [email for email in emails if
                                             email.strip() != '' and not email_valido(email)]
                            if nenblanco or ienblanco:
                                return JsonResponse({"result": "bad",
                                                     "mensaje": u"Los datos de los integrantes del comité científico evaluador son obligatorios a excepción del e-mail"})

                            if emailnovalido:
                                return JsonResponse({"result": "bad",
                                                     "mensaje": u"El formato de una o varias direcciones de e-mail de los integrantes del comité científico evaluador no es válido"})

                            listado = [nombre.strip() for nombre in integrantes if nombre.strip() != '']
                            repetido = [x for x, y in collections.Counter(listado).items() if y > 1]
                            if repetido:
                                return JsonResponse({"result": "bad",
                                                     "mensaje": u"El integrante [%s] del comité científico evaluador está repetido" % (
                                                     repetido[0])})

                        solicitud.nombre = f.cleaned_data['nombre']
                        solicitud.motivo = f.cleaned_data['motivo']
                        solicitud.evento = f.cleaned_data['evento']
                        solicitud.pais = f.cleaned_data['pais']
                        solicitud.ciudad = f.cleaned_data['ciudad']
                        solicitud.fecharecepcion = f.cleaned_data['fecharecepcion']
                        solicitud.fechaaprobacion = f.cleaned_data['fechaaprobacion']
                        solicitud.fechapublicacion = f.cleaned_data['fechapublicacionponencia'] if int(
                            f.cleaned_data['estadopublicacionponencia']) == 1 else None
                        solicitud.enlace = f.cleaned_data['enlace']
                        solicitud.areaconocimiento = f.cleaned_data['areaconocimiento']
                        solicitud.subareaconocimiento = f.cleaned_data['subareaconocimiento']
                        solicitud.subareaespecificaconocimiento = f.cleaned_data['subareaespecificaconocimiento']
                        solicitud.lineainvestigacion = f.cleaned_data['lineainvestigacion']
                        solicitud.sublineainvestigacion = f.cleaned_data['sublineainvestigacion']
                        solicitud.provieneproyecto = f.cleaned_data['provieneproyecto']
                        solicitud.tipoproyecto = f.cleaned_data['tipoproyecto'] if f.cleaned_data[
                            'provieneproyecto'] else None
                        solicitud.comitecientifico = f.cleaned_data['comitecientifico']
                        solicitud.estadopublicacion = f.cleaned_data['estadopublicacionponencia']
                        solicitud.observacion = ""
                        solicitud.estado = estadosolicitud

                        if f.cleaned_data['provieneproyecto']:
                            if int(f.cleaned_data['tipoproyecto']) != 3:
                                solicitud.proyectointerno = f.cleaned_data['proyectointerno']
                                solicitud.proyectoexterno = None
                            else:
                                solicitud.proyectoexterno = f.cleaned_data['proyectoexterno']
                                solicitud.proyectointerno = None
                        else:
                            solicitud.proyectoexterno = None
                            solicitud.proyectointerno = None

                        # if solicitud.tiposolicitud == 2 or solicitud.tiposolicitud == 3:
                        #     solicitud.motivo = f.cleaned_data['motivo']
                        #     solicitud.nombre = f.cleaned_data['nombre']
                        #     solicitud.fechapublicacion = f.cleaned_data['fechapublicacion']
                        #     solicitud.areaconocimiento = f.cleaned_data['areaconocimiento']
                        #     solicitud.subareaconocimiento = f.cleaned_data['subareaconocimiento']
                        #     solicitud.subareaespecificaconocimiento = f.cleaned_data['subareaespecificaconocimiento']
                        #     if solicitud.tiposolicitud == 2:
                        #         solicitud.evento = f.cleaned_data['evento']
                        #         solicitud.enlace = f.cleaned_data['enlace']
                        # else:
                        #     solicitud.motivo = f.cleaned_data['motivo']
                        #
                    else:
                        solicitud.observacion = ""
                        solicitud.estado = estadosolicitud
                    # solicitud.save(request)

                    if 'archivo' in request.FILES:
                        newfile = request.FILES['archivo']
                        newfile._name = generar_nombre("solicitudpublicacion_", newfile._name)
                        solicitud.archivo = newfile

                    if 'archivocertificado' in request.FILES:
                        newfile = request.FILES['archivocertificado']
                        newfile._name = generar_nombre("solicitudpublicacioncertificacion_", newfile._name)
                        solicitud.archivocertificado = newfile

                    if 'archivoparticipacion' in request.FILES:
                        newfile = request.FILES['archivoparticipacion']
                        newfile._name = generar_nombre("solicitudpublicacionparticipacion_", newfile._name)
                        solicitud.archivoparticipacion = newfile

                    if 'archivocomite' in request.FILES:
                        newfile = request.FILES['archivocomite']
                        newfile._name = generar_nombre("solicitudpublicacioncomite_", newfile._name)
                        solicitud.archivocomite = newfile

                    if solicitud.tiposolicitud == 2:
                        integrantes = request.POST.getlist('integrante[]')
                        instituciones = request.POST.getlist('institucion[]')
                        emails = request.POST.getlist('email[]')

                        comiteevaluador = "|".join([
                            integrante.upper().strip() + "," + institucion.upper().strip() + "," + email.lower().strip()
                            for integrante, institucion, email in zip(integrantes, instituciones, emails)])
                        solicitud.integrantecomite = comiteevaluador

                    solicitud.save(request)

                    log(u'Modifico solicitud: %s' % solicitud, request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                    errorformulario = f._errors
                    raise NameError('Error')
            except Exception as ex:
                msg = ex.__str__()
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos. Detalle: %s" % (msg)})

        elif action == 'editsolicitudlibro':
            try:
                f = SolicitudPublicacionLibroForm(request.POST, request.FILES)
                archivocarta = archivopublicacion = None

                if 'archivocertificado' in request.FILES:
                    archivocarta = request.FILES['archivocertificado']
                    descripcionarchivo = 'Archivo carta de aceptación'

                    # Validar el archivo
                    resp = validar_archivo(descripcionarchivo, archivocarta, ['PDF', 'DOC', 'DOCX'], '10MB')
                    if resp['estado'] != "OK":
                        return JsonResponse({"result": "bad", "mensaje": resp["mensaje"]})

                if 'archivo' in request.FILES:
                    archivopublicacion = request.FILES['archivo']
                    descripcionarchivo = 'Archivo publicación'

                    # Validar el archivo
                    resp = validar_archivo(descripcionarchivo, archivopublicacion, ['PDF', 'DOC', 'DOCX'], '10MB')
                    if resp['estado'] != "OK":
                        return JsonResponse({"result": "bad", "mensaje": resp["mensaje"]})

                if f.is_valid():
                    solicitud = SolicitudPublicacion.objects.get(pk=request.POST['id'])
                    estadosolicitud = obtener_estado_solicitud(8, 1)  # SOLICITADO

                    solicitud.nombre = f.cleaned_data['titulo']
                    solicitud.fechapublicacion = f.cleaned_data['fechapublicacion']
                    solicitud.areaconocimiento = f.cleaned_data['areaconocimiento']
                    solicitud.subareaconocimiento = f.cleaned_data['subareaconocimiento']
                    solicitud.subareaespecificaconocimiento = f.cleaned_data['subareaespecificaconocimiento']
                    solicitud.observacion = ""
                    solicitud.estado = estadosolicitud

                    if archivocarta:
                        archivocarta._name = generar_nombre("solicitudpublicacioncertificacion_", archivocarta._name)
                        solicitud.archivocertificado = archivocarta

                    if archivopublicacion:
                        archivopublicacion._name = generar_nombre("solicitudpublicacion_", archivopublicacion._name)
                        solicitud.archivo = archivopublicacion

                    solicitud.save(request)

                    log(u'Modifico solicitud de publicación de libro: %s' % solicitud, request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                    errorformulario = f._errors
                    raise NameError('Error')
            except Exception as ex:
                msg = ex.__str__()
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos. Detalle: %s" % (msg)})

        elif action == 'editsolicitudcapitulo':
            try:
                f = CapituloLibroHojaVidaForm(request.POST, request.FILES)

                archivolibro = archivocapitulo = archivocarta = archivopares = None

                if 'archivolibro' in request.FILES:
                    archivolibro = request.FILES['archivolibro']
                    descripcionarchivo = 'Libro'
                    resp = validar_archivo(descripcionarchivo, archivolibro, ['pdf', 'doc', 'docx'], '10MB')
                    if resp['estado'] != "OK":
                        return JsonResponse(
                            {"result": "bad", "titulo": "Atención!!!", "mensaje": resp["mensaje"], "showSwal": "True",
                             "swalType": "warning"})

                if 'archivocapitulo' in request.FILES:
                    archivocapitulo = request.FILES['archivocapitulo']
                    descripcionarchivo = 'Capítulo de libro'
                    resp = validar_archivo(descripcionarchivo, archivocapitulo, ['pdf', 'doc', 'docx'], '10MB')
                    if resp['estado'] != "OK":
                        return JsonResponse(
                            {"result": "bad", "titulo": "Atención!!!", "mensaje": resp["mensaje"], "showSwal": "True",
                             "swalType": "warning"})

                if 'archivocarta' in request.FILES:
                    archivocarta = request.FILES['archivocarta']
                    descripcionarchivo = 'Carta de aceptación'
                    resp = validar_archivo(descripcionarchivo, archivocarta, ['pdf', 'doc', 'docx'], '10MB')
                    if resp['estado'] != "OK":
                        return JsonResponse(
                            {"result": "bad", "titulo": "Atención!!!", "mensaje": resp["mensaje"], "showSwal": "True",
                             "swalType": "warning"})

                if 'archivopares' in request.FILES:
                    archivopares = request.FILES['archivopares']
                    descripcionarchivo = 'Informe revsión pares'
                    resp = validar_archivo(descripcionarchivo, archivopares, ['pdf', 'doc', 'docx'], '10MB')
                    if resp['estado'] != "OK":
                        return JsonResponse(
                            {"result": "bad", "titulo": "Atención!!!", "mensaje": resp["mensaje"], "showSwal": "True",
                             "swalType": "warning"})

                if f.is_valid():
                    solicitud = SolicitudPublicacion.objects.get(pk=request.POST['id'])
                    estadosolicitud = obtener_estado_solicitud(8, 1)  # SOLICITADO

                    solicitud.nombre = f.cleaned_data['titulocapitulo']
                    solicitud.motivo = f.cleaned_data['resumen']
                    solicitud.evento = f.cleaned_data['titulolibro']
                    solicitud.codigoisbn = f.cleaned_data['codigoisbn']
                    solicitud.paginas = f.cleaned_data['paginas']
                    solicitud.editorcompilador = f.cleaned_data['editorcompilador']
                    solicitud.fechapublicacion = f.cleaned_data['fechapublicacion']
                    solicitud.filiacion = f.cleaned_data['filiacion']
                    solicitud.areaconocimiento = f.cleaned_data['areaconocimiento']
                    solicitud.subareaconocimiento = f.cleaned_data['subareaconocimiento']
                    solicitud.subareaespecificaconocimiento = f.cleaned_data['subareaespecificaconocimiento']
                    solicitud.observacion = ""
                    solicitud.estado = estadosolicitud

                    if archivolibro:
                        archivolibrog = archivolibro
                        archivolibrog._name = generar_nombre("libro", archivolibrog._name)
                        solicitud.archivocertificado = archivolibrog

                    if archivocapitulo:
                        archivocapitulog = archivocapitulo
                        archivocapitulog._name = generar_nombre("capitulolibro", archivocapitulog._name)
                        solicitud.archivo = archivocapitulog

                    if archivocarta:
                        archivocartag = archivocarta
                        archivocartag._name = generar_nombre("cartaaceptacion", archivocartag._name)
                        solicitud.archivoparticipacion = archivocartag

                    if archivopares:
                        archivoparesg = archivopares
                        archivoparesg._name = generar_nombre("revisionpares", archivoparesg._name)
                        solicitud.archivocomite = archivoparesg

                    solicitud.save(request)

                    log(u'%s modificó solicitud de publicación de capítulo de libro: %s' % (persona, solicitud),
                        request, "edit")
                    return JsonResponse(
                        {"result": "ok", "titulo": "Proceso exitoso!!!", "mensaje": u"Registro actualizado con éxito",
                         "showSwal": True})
                else:
                    errorformulario = f._errors
                    raise NameError('Error en el formulario')
            except Exception as ex:
                msg = ex.__str__()
                transaction.set_rollback(True)
                return JsonResponse(
                    {"result": "bad", "titulo": "Error", "mensaje": u"Error al guardar los datos. [%s]" % msg,
                     "showSwal": "True", "swalType": "error"})

        elif action == 'addcabsolicitud_docente':
            try:
                if Departamento.objects.filter(integrantes__isnull=False, responsable=persona,
                                               status=True).distinct().exists():
                    cabecera = CapCabeceraSolicitudDocente(capeventoperiodo_id=int(request.POST['id']),
                                                           solicita=persona,
                                                           fechasolicitud=datetime.now().date(),
                                                           estadosolicitud=variable_valor('PENDIENTE_CAPACITACION'),
                                                           participante=persona)
                    cabecera.save(request)
                    log(u'Ingreso Cabecera Solicitud de Evento Docente: %s' % cabecera, request, "add")
                    detalle = CapDetalleSolicitudDocente(cabecera=cabecera,
                                                         aprueba=persona,
                                                         observacion="SOLICITUD DE INSCRIPCIÓN",
                                                         fechaaprobacion=datetime.now().date(),
                                                         estado=2)
                    detalle.save(request)
                    detalle.mail_notificar_talento_humano(request.session['nombresistema'], True)
                    log(u'Ingreso Detalle Solicitud de Evento Docente: %s' % detalle, request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                    cabecera = CapCabeceraSolicitudDocente(capeventoperiodo_id=int(request.POST['id']),
                                                           solicita=persona,
                                                           fechasolicitud=datetime.now().date(),
                                                           estadosolicitud=variable_valor('SOLICITUD_CAPACITACION'),
                                                           participante=persona)
                    cabecera.save(request)
                    log(u'Ingreso Cabecera Solicitud de Evento Docente: %s' % cabecera, request, "add")
                    detalle = CapDetalleSolicitudDocente(cabecera=cabecera,
                                                         aprueba=persona,
                                                         observacion="SOLICITUD DE INSCRIPCIÓN",
                                                         fechaaprobacion=datetime.now().date(),
                                                         estado=1)
                    detalle.save(request)
                    detalle.mail_notificar_jefe_departamento(request.session['nombresistema'], False)
                    log(u'Ingreso Detalle Solicitud de Evento Docente: %s' % detalle, request, "add")
                    return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        elif action == 'deleventoaprobacion_docente':
            try:
                cabecera = CapCabeceraSolicitudDocente.objects.get(pk=int(request.POST['id']))
                cabecera.capdetallesolicituddocente_set.all().delete()
                cabecera.delete()
                log(u'Elimino Cabecera Solicitud de Evento Docente: %s' % cabecera, request, "del")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al elminar los datos."})

        elif action == 'meses_anio':
            try:
                distributivo = DistributivoPersona.objects.get(pk=int(request.POST['id']))
                anio = request.POST['anio']
                lista = []
                for elemento in TrabajadorDiaJornada.objects.filter(persona=distributivo.persona, anio=anio).order_by(
                        'mes').distinct():
                    if [elemento.mes, elemento.rep_mes()] not in lista:
                        lista.append([elemento.mes, elemento.rep_mes()])
                return JsonResponse({"result": "ok", "lista": lista})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'meses_anio_log':
            try:
                # distributivo = DistributivoPersona.objects.get(pk=int(request.POST['id']))
                personad = Persona.objects.get(pk=int(request.POST['id']))
                anio = request.POST['anio']
                lista = []
                for e in LogDia.objects.filter(persona=personad, fecha__year=anio, status=True).order_by(
                        'fecha').distinct():
                    if [e.fecha.month, MESES_CHOICES[e.fecha.month - 1][1]] not in lista:
                        lista.append([e.fecha.month, MESES_CHOICES[e.fecha.month - 1][1]])
                return JsonResponse({"result": "ok", "lista": lista})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'addarchivo':
            try:
                form = DocumentoInscripcionForm(request.POST, request.FILES)
                inscripcion = Inscripcion.objects.get(pk=request.POST['id'])
                arch = request.FILES['archivo']
                extencion = arch._name.split('.')
                exte = extencion[1]
                if arch.size > 5242880:
                    return JsonResponse({"result": "bad", "mensaje": u"Error, archivo mayor a 5 Mb."})
                if not exte == 'pdf':
                    return JsonResponse({"result": "bad", "mensaje": u"Solo archivos .pdf"})
                if form.is_valid():
                    nfile = request.FILES['archivo']
                    nfile._name = generar_nombre("documentos_", nfile._name)
                    archivo = Archivo(fecha=datetime.now().date(),
                                      archivo=nfile,
                                      tipo_id=int(request.POST['tipo']),
                                      inscripcion=inscripcion)
                    archivo.save(request)
                    log(u'Adiciono documento de a hoja de vida: %s - %s' % (inscripcion, archivo), request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'editarchivo':
            try:
                form = DocumentoInscripcionForm(request.POST, request.FILES)
                archivo = Archivo.objects.get(pk=request.POST['id'])
                if form.is_valid():
                    if 'archivo' in request.FILES:
                        nfile = request.FILES['archivo']
                        nfile._name = generar_nombre("documentos_", nfile._name)
                        archivo.fecha = datetime.now().date()
                        archivo.tipo_id = int(request.POST['tipo'])
                        archivo.archivo = nfile
                    else:
                        archivo.fecha = datetime.now().date()
                        archivo.tipo_id = int(request.POST['tipo'])
                    archivo.save(request)
                    log(u'Edito documento de a hoja de vida: %s - %s' % (perfilprincipal.inscripcion, archivo), request,
                        "edit")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'delarchivo':
            try:
                archivo = Archivo.objects.get(pk=request.POST['id'])
                archivo.status = False
                archivo.save(request)
                log(u'Edito documento de a hoja de vida: %s - %s' % (perfilprincipal.inscripcion, archivo), request,
                    "del")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'editarticulos':
            try:
                articulos = ArticuloInvestigacion.objects.get(pk=request.POST['id'])
                f = ArticuloInvestigacionForm(request.POST)
                if f.is_valid():
                    articulos.areaconocimiento = f.cleaned_data['areaconocimiento']
                    articulos.subareaconocimiento = f.cleaned_data['subareaconocimiento']
                    articulos.subareaespecificaconocimiento = f.cleaned_data['subareaespecificaconocimiento']
                    articulos.save(request)
                    log(u"Edito articulo hoja de vida: %s" % articulos, request, "edit")
                    return JsonResponse({'result': 'ok'})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', 'mensaje': u'Error: al editar'})

        elif action == 'solicitudjusti':
            try:
                f = SolicitudJustificacionMarcadaForm(request.POST, request.FILES)
                archivo = None
                if 'archivo' in request.FILES:
                    archivo = request.FILES['archivo']
                    archivo._name = remover_caracteres_especiales_unicode(archivo._name)
                    ext = archivo._name.split('.')
                    size = archivo.size
                    if not ext[len(ext) - 1] in ['pdf', 'jpg', 'jpeg', 'png']:
                        raise NameError(
                            'No se permite el archivo .%s. Archivos permitidos .pdf, .jpg, .png' % (ext[len(ext) - 1]))
                    if not size < 4194304:
                        raise NameError('Tamaño de archivo mayor a 4mb')
                if not f.is_valid():
                    raise NameError('Datos incorrectos.')
                solicitud = SolicitudJustificacionMarcada(
                    solicita=persona,
                    estado=1,
                    observacion=f.cleaned_data['observacion'],
                    fecha=datetime.now(),
                    documento=archivo
                )
                solicitud.save(request)
                log('Agrego una solicitud de justificacion de marcadas %s' % (solicitud.__str__()), request, 'add')
                if not 'lista_items1' in request.POST:
                    raise NameError('Debe adicionar un registro de justificacion')
                for diassolicitud in json.loads(request.POST['lista_items1']):
                    fecha = datetime.strptime(diassolicitud['fecha'], '%Y-%m-%d').date()
                    if persona.logdia_set.filter(fecha=fecha, status=True).exists():
                        logdia = persona.logdia_set.filter(fecha=fecha, status=True)[0]
                        if logdia.persona.historialjornadatrabajador_set.filter(fechainicio__lte=logdia.fecha,
                                                                                fechafin__gte=logdia.fecha).exists():
                            logdia.jornada = \
                            logdia.persona.historialjornadatrabajador_set.filter(fechainicio__lte=logdia.fecha,
                                                                                 fechafin__gte=logdia.fecha).order_by(
                                'fechainicio')[0].jornada
                        elif logdia.persona.historialjornadatrabajador_set.filter(fechainicio__lte=logdia.fecha,
                                                                                  fechafin=None).exists():
                            logdia.jornada = \
                            logdia.persona.historialjornadatrabajador_set.filter(fechainicio__lte=logdia.fecha,
                                                                                 fechafin=None)[0].jornada
                    else:
                        logdia = LogDia(persona=persona,
                                        fecha=fecha,
                                        cantidadmarcadas=1)
                        if persona.historialjornadatrabajador_set.filter(fechainicio__lte=fecha,
                                                                         fechafin__gte=fecha).exists():
                            logdia.jornada = persona.historialjornadatrabajador_set.filter(fechainicio__lte=fecha,
                                                                                           fechafin__gte=fecha).order_by(
                                'fechainicio')[0].jornada
                        elif persona.historialjornadatrabajador_set.filter(fechainicio__lte=fecha,
                                                                           fechafin=None).exists():
                            logdia.jornada = \
                            persona.historialjornadatrabajador_set.filter(fechainicio__lte=fecha, fechafin=None)[
                                0].jornada
                    logdia.save(request)
                    time = datetime.strptime(diassolicitud['fecha'] + ' ' + diassolicitud['hora'], '%Y-%m-%d %H:%M')
                    detallesolicitud = DetalleSolicitudJustificacionMarcada(
                        solicitud=solicitud,
                        dia=logdia,
                        secuencia=diassolicitud['secuencia'],
                        tiposolcitud=diassolicitud['soli'],
                        hora=time,
                    )
                    if diassolicitud['marcada'] != '':
                        if logdia.logmarcada_set.filter(pk=int(diassolicitud['marcada'])).exists():
                            logmarcada = logdia.logmarcada_set.filter(pk=int(diassolicitud['marcada']))[0]
                            detallesolicitud.marcada = logmarcada
                    else:
                        if logdia.logmarcada_set.filter(time=time).exists():
                            logmarcada = logdia.logmarcada_set.filter(time=time)[0]
                            detallesolicitud.marcada = logmarcada
                    detallesolicitud.save(request)
                    log("Agrego detalle de justificacion de marcada %s" % (detallesolicitud.__str__()), request, 'add')
                historial = HistorialSolicitudJustificacionMarcada(
                    solicitud=solicitud,
                    observacion=f.cleaned_data['observacion'],
                    estado=1,
                    fecha=datetime.now(),
                    persona=persona,
                )
                historial.save(request)
                log("Agrego historial de justificacion de marcada %s" % (detallesolicitud.__str__()), request, 'add')
                if not solicitud.solicita.departamento_set.filter(status=True).exists():
                    raise NameError('No tiene asignado departamento')
                liderdepartamento = solicitud.solicita.departamento_set.filter(status=True)[0].responsable
                noti(
                    'Solicitud justificación de marcadas',
                    f'El servidor {solicitud.solicita.__str__().title()} ha realizado una solicitud justificación de marcadas',
                    liderdepartamento,
                    None,
                    'th_aprobarpermiso?action=justifiacionmarcadas',
                    solicitud.id,
                    2,
                    'sagest',
                    SolicitudJustificacionMarcada,
                    request
                )
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse(
                    {"result": "bad", "mensaje": u"Error al guardar los datos. Detalle: %s" % (ex.__str__())})

        elif action == 'informemarcada':
            mensaje = "Problemas al generar el informe de marcadas."
            try:
                fini = convertir_fecha_invertida(request.POST['fini'])
                ffin = convertir_fecha_invertida(request.POST['ffin'])
                data['personaadmin'] = persona = request.session['persona']
                marcadas = LogDia.objects.filter(persona=persona, fecha__gte=fini, fecha__lte=ffin,
                                                 status=True).order_by('fecha')
                return conviert_html_to_pdf('th_hojavida/informemarcadas.html',
                                            {'pagesize': 'A4',
                                             'marcadas': marcadas,
                                             'persona': persona, 'fechainicio': fini, 'fechafin': ffin,
                                             'hoy': datetime.now().date()
                                             })
            except Exception as ex:
                return HttpResponseRedirect("/th_hojavida?info=%s" % mensaje)

        elif action == 'mostrarseleccionadas':
            try:
                fecha = datetime.now().date()
                data['periodo'] = periodo = request.session['periodoadmision']
                if periodo.preferenciaadmisionfechatope:
                    if (periodo.preferenciaadmisionfechatope >= fecha):
                        data['accesopreferencia'] = True
                    else:
                        data['accesopreferencia'] = False
                else:
                    data['accesopreferencia'] = False
                data['asignaturaspreferencias'] = materiaspreferencias = AsignaturaMallaPreferencia.objects.filter(
                    persona=persona, periodo=periodo, status=True)
                data['totalpreferencia'] = materiaspreferencias.count()
                template = get_template("th_hojavida/mostrarseleccionadas.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'malla': json_content})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": "Error al guardar los datos."})

        elif action == 'afinidad_malla':
            try:
                # iditem
                if 'idp' in request.POST:
                    asignaturamallapreferencia = AsignaturaMallaPreferencia.objects.get(pk=request.POST['idp'])
                    data['materia'] = asignaturamallapreferencia.asignaturamalla
                else:
                    asignaturamalla = AsignaturaMalla.objects.get(pk=request.POST['iditem'])
                    data['materia'] = asignaturamalla
                data['titulacion'] = Titulacion.objects.filter(persona=persona, status=True,
                                                               titulo__nivel__id=4).exclude(
                    titulo__grado__id=3).order_by('titulo__grado__id')
                data['title'] = u'Detalle Título'
                template = get_template("th_hojavida/afinidad.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'data': json_content})
            except Exception as ex:
                return JsonResponse({"result": "bad"})

        elif action == 'afinidad_publicaciones':
            try:
                # iditem
                asignaturamallapreferencia = AsignaturaMallaPreferencia.objects.get(pk=request.POST['idp'])
                data['materia'] = asignaturamallapreferencia.asignaturamalla
                data['titulacion'] = ArticuloInvestigacion.objects.select_related().filter(
                    participantesarticulos__profesor=persona, status=True,
                    participantesarticulos__status=True).order_by('-fechapublicacion')
                data['title'] = u'Detalle de Publicaciones de investifación'
                template = get_template("th_hojavida/afinidadpublicaciones.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'data': json_content})
            except Exception as ex:
                return JsonResponse({"result": "bad"})

        elif action == 'afirmaasignaturapreferencia':
            try:
                materia = AsignaturaMalla.objects.get(pk=int(encrypt(request.POST['id'])), status=True)
                nombres = materia.asignatura.nombre
                return JsonResponse({"result": "ok", 'asignatura': nombres, 'idasigmalla': materia.id})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'conasignaturapreferencia':
            try:
                materia = AsignaturaMallaPreferencia.objects.get(pk=request.POST['id'], status=True)
                nombres = materia.asignaturamalla.asignatura.nombre
                return JsonResponse({"result": "ok", 'asignatura': nombres, 'idmateripreferencia': materia.id})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'addasignaturapreferencia':
            try:
                mallapreferencia = AsignaturaMallaPreferencia(persona_id=request.POST['idpersona'],
                                                              asignaturamalla_id=request.POST['idasignaturamalla'],
                                                              periodo_id=request.POST['idperiodo'])
                mallapreferencia.save(request)
                log(u'Adiciono Asignatura de preferencia: %s' % mallapreferencia, request, "add")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'delasignaturapreferencia':
            try:
                asigpreferencia = AsignaturaMallaPreferencia.objects.get(pk=request.POST['idmatpreferencia'])
                asigpreferencia.status = False
                asigpreferencia.save(request)
                log(u'Elimino Asignatura de preferencia: %s' % asigpreferencia, request, "add")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        elif action == 'silabopdf':
            try:
                silabo = Silabo.objects.get(pk=request.POST['id'], status=True)
                return conviert_html_to_pdf(
                    'pro_planificacion/silabo_pdf.html',
                    {
                        'pagesize': 'A4',
                        'data': silabo.silabo_pdf(),
                    }
                )
            except Exception as ex:
                pass

        elif action == 'mostrarmalla':
            try:
                data['admisionperiodo'] = periodo = request.session['periodoadmision']
                data['malla'] = malla = Malla.objects.get(pk=int(request.POST['id']))
                materiaspreferencias = AsignaturaMallaPreferencia.objects.values_list('asignaturamalla').filter(
                    persona=persona, periodo=periodo, status=True)
                data['nivelesdemallas'] = NivelMalla.objects.filter(status=True,
                                                                    pk__in=AsignaturaMalla.objects.values_list(
                                                                        'nivelmalla').filter(malla=malla).exclude(
                                                                        pk__in=materiaspreferencias)).order_by('id')
                data['asignaturasmallas'] = AsignaturaMalla.objects.filter(malla=malla).exclude(
                    pk__in=materiaspreferencias)
                template = get_template("th_hojavida/mostrarmalla.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'malla': json_content})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": "Error al guardar los datos."})

        elif action == 'addbitacoraposgrado':
            try:
                persona = request.session['persona']
                form = BitacoraForm(request.POST)
                if persona.contratodip_set.filter(status=True).exists():
                    fecha = datetime.today()
                    fecbitaco = convertir_fecha_hora(u"%s %s" % (request.POST['fecha'], request.POST['hora']))
                    if datetime.fromisoformat('2022-07-01').date() != fecbitaco.date():
                        if (fecha - fecbitaco).days > 2:
                            return JsonResponse(
                                {'result': 'bad', 'mensaje': "Excedió fecha limite para subir bitácora "})
                newfile = None
                if 'archivo' in request.FILES:
                    newfile = request.FILES['archivo']
                    if newfile:
                        newfilesd = newfile._name
                        ext = newfilesd[newfilesd.rfind("."):]
                        if newfile.size > 12582912:
                            return JsonResponse({"result": "bad", "mensaje": u"Error, archivo mayor a 12 Mb."})
                        if newfile:
                            newfile._name = generar_nombre("bitacora", newfile._name)
                if form.is_valid():
                    bitacora = BitacoraActividadDiaria(persona=persona,
                                                       fecha=convertir_fecha_hora(
                                                           u"%s %s" % (request.POST['fecha'], request.POST['hora'])),
                                                       departamento=persona.mi_departamento(),
                                                       descripcion=form.cleaned_data['descripcion'],
                                                       link=form.cleaned_data['link'],
                                                       tiposistema=form.cleaned_data['tiposistema'],
                                                       archivo=newfile,
                                                       departamento_requiriente=form.cleaned_data[
                                                           'departamento_requiriente']
                                                       )
                    bitacora.actividades = form.cleaned_data['actividades']
                    bitacora.save(request)
                    log(u'Adicionó bitacora de actividades diarias %s' % (bitacora), request, "add")
                    return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'editbitacoraposgrado':
            try:
                persona = request.session['persona']
                form = BitacoraForm(request.POST)
                newfile = None
                if 'archivo' in request.FILES:
                    newfile = request.FILES['archivo']
                    if newfile:
                        newfilesd = newfile._name
                        ext = newfilesd[newfilesd.rfind("."):]
                        if newfile.size > 12582912:
                            return JsonResponse({"result": "bad", "mensaje": u"Error, archivo mayor a 12 Mb."})
                        if newfile:
                            newfile._name = generar_nombre("bitacora", newfile._name)
                if form.is_valid():
                    bitacora = BitacoraActividadDiaria.objects.get(pk=request.POST['id'])
                    bitacora.actividades = form.cleaned_data['actividades']
                    bitacora.fecha = convertir_fecha_hora(u"%s %s" % (request.POST['fecha'], request.POST['hora']))
                    bitacora.departamento = persona.mi_departamento()
                    bitacora.descripcion = form.cleaned_data['descripcion']
                    bitacora.link = form.cleaned_data['link']
                    bitacora.tiposistema = form.cleaned_data['tiposistema']
                    bitacora.departamento_requiriente = form.cleaned_data['departamento_requiriente']
                    if newfile:
                        bitacora.archivo = newfile
                    bitacora.save(request)
                    log(u'Editó bitacora de actividad diaria %s' % (bitacora), request, "edit")
                    return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'deleteinftecnico':
            try:
                with transaction.atomic():
                    instancia = InformeTecnico.objects.get(pk=int(request.POST['id']))
                    instancia.status = False
                    instancia.save(request)
                    log(u'Elimino informe tecnico: #%s %s' % (instancia, instancia.contrato.persona), request, "delete")
                    res_json = {"error": False}
            except Exception as ex:
                res_json = {'error': True, "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)

        elif action == 'cargaradicionartitulo':
            try:

                data['form'] = TituloHojaVidaForm()

                template = get_template('th_hojavida/addtitulo.html')
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'html': json_content})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al cargar los datos."})

        elif action == 'cargaradicionartituloaspirante':
            try:

                data['form2'] = TituloHojaVidaAdmisionForm()

                template = get_template('th_hojavida/addtituloaspirante.html')
                return JsonResponse({"result": True, 'data': template.render(data)})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al cargar los datos."})

        elif action == 'addTitulo':
            try:
                f = TituloHojaVidaForm(request.POST)
                if f.is_valid():
                    if Titulo.objects.filter(nombre__unaccent=f.cleaned_data['nombre'],
                                             nivel=f.cleaned_data['nivel']).exists():
                        return JsonResponse({"result": "bad", "mensaje": u"El registro ya existe."})
                    titulo = Titulo(nombre=f.cleaned_data['nombre'],
                                    abreviatura=f.cleaned_data['abreviatura'],
                                    nivel=f.cleaned_data['nivel'],
                                    grado=f.cleaned_data['grado'],
                                    )
                    titulo.save(request)
                    log(u'Adiciono nuevo titulo: %s' % titulo, request, "add")
                    messages.success(request, 'Se guado exitosamente')
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'editmarcada':
            try:
                form = CambiarMarcadaForm(request.POST)
                if form.is_valid():
                    if puede_realizar_accion(request, 'sagest.puede_agregar_marcada_log'):
                        logmarcada = LogMarcada.objects.get(id=int(request.POST['id']))
                        time1 = form.cleaned_data['hora']
                        time = datetime(logmarcada.time.year, logmarcada.time.month, logmarcada.time.day,
                                        time1.hour, time1.minute, time1.second)
                        logmarcada.time = time
                        logmarcada.logdia.procesado = False
                        logmarcada.logdia.save(request)
                        logmarcada.save(request)
                        cm = logmarcada.logdia.logmarcada_set.filter(status=True).count()
                        logmarcada.logdia.cantidadmarcadas = cm
                        if logmarcada.logdia.persona.marcadasdia_set.filter(fecha=logmarcada.logdia.fecha).exists():
                            marcadadia = \
                                logmarcada.logdia.persona.marcadasdia_set.filter(fecha=logmarcada.logdia.fecha,
                                                                                 status=True)[0]
                            if marcadadia.registromarcada_set.all().exists():
                                registrosm = marcadadia.registromarcada_set.all()
                                for r in registrosm:
                                    r.delete()

                        else:
                            marcadadia = MarcadasDia(persona=logmarcada.logdia.persona,
                                                     fecha=logmarcada.logdia.fecha,
                                                     logdia=logmarcada.logdia,
                                                     segundos=0)
                            marcadadia.save(request)
                        if (cm % 2) == 0:
                            marini = 1
                            for dl in logmarcada.logdia.logmarcada_set.filter(status=True).order_by("time"):
                                if marini == 2:
                                    salida = dl.time
                                    marini = 1
                                    registro = None
                                    if marcadadia.registromarcada_set.filter(status=True, entrada=entrada).exists():
                                        registro = \
                                            marcadadia.registromarcada_set.filter(status=True, entrada=entrada)[0]
                                        registro.entrada = entrada
                                        registro.salida = salida
                                        registro.segundos = (salida - entrada).seconds

                                    else:
                                        registro = RegistroMarcada(marcada=marcadadia,
                                                                   entrada=entrada,
                                                                   salida=salida,
                                                                   segundos=(salida - entrada).seconds)
                                    registro.save(request)

                                    marcadadia.actualizar_marcadas()
                                else:
                                    entrada = dl.time
                                    marini += 1
                            logmarcada.logdia.procesado = True
                        else:
                            logmarcada.logdia.cantidadmarcadas = 0
                            logmarcada.logdia.procesado = False
                        logmarcada.logdia.save(request)
                        calculando_marcadas(request, logmarcada.time.date(), logmarcada.time.date(),
                                            logmarcada.logdia.persona)

                        log(u'editó marcada : %s' % logmarcada, request, "edit")

                        return JsonResponse({"result": False}, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Error al guardar los datos."}, safe=False)

        elif action == 'cargar_campoamplio':
            try:
                t = Titulo.objects.filter(pk=request.POST['id']).first()
                lista = []
                campotitulo = CamposTitulosPostulacion.objects.filter(status=True, titulo=t).first()
                if campotitulo:
                    for ca in campotitulo.campoamplio.all():
                        lista.append([ca.id, ca.__str__()])
                else:
                    for ca in AreaConocimientoTitulacion.objects.all():
                        lista.append([ca.id, ca.__str__()])
                return JsonResponse({'result': 'ok', 'lista': lista})
            except Exception as ex:
                return JsonResponse(
                    {"result": "bad", "mensaje": u"Seleccione título: mínimo de tercer nivel o superior."})

        elif action == 'cargar_todos':
            try:
                t = Titulo.objects.filter(pk=request.POST['idtitulo']).first()
                listaca = []
                listace = []
                listacd = []
                campotitulo, campoamplio, campoespecifico, campodetallado = None, None, None, None
                if CamposTitulosPostulacion.objects.filter(status=True, titulo=t).exists():
                    campotitulo = CamposTitulosPostulacion.objects.filter(status=True, titulo=t).first()
                    campoamplio = AreaConocimientoTitulacion.objects.filter(status=True,
                                                                            id__in=campotitulo.campoamplio.all().values_list(
                                                                                'id', flat=True))
                    if campoamplio:
                        for ca in campoamplio:
                            listaca.append([ca.id, ca.__str__()])
                    campoespecifico = SubAreaConocimientoTitulacion.objects.filter(status=True,
                                                                                   id__in=campotitulo.campoespecifico.all().values_list(
                                                                                       'id', flat=True))
                    if campoespecifico:
                        for ce in campoespecifico:
                            listace.append([ce.id, ce.__str__()])
                    campodetallado = SubAreaEspecificaConocimientoTitulacion.objects.filter(status=True,
                                                                                            id__in=campotitulo.campodetallado.all().values_list(
                                                                                                'id', flat=True))
                    if campodetallado:
                        for cd in campodetallado:
                            listacd.append([cd.id, cd.__str__()])
                    return JsonResponse({"result": "ok", "campoamplio": listaca, "campoespecifico": listace,
                                         "campodetallado": listacd})
                else:
                    return JsonResponse({"result": "no"})
            except Exception as ex:
                return JsonResponse(
                    {"result": "bad", "mensaje": u"Seleccione título: mínimo de tercer nivel o superior."})

        elif action == 'addarchivoevidencias':
            try:
                form = ArchivoInformesForm(request.POST, request.FILES)
                if form.is_valid():
                    newfile = request.FILES['archivo']
                    extension = newfile._name.split('.')
                    tam = len(extension)
                    exte = extension[tam - 1]
                    if newfile.size > 100194304:
                        transaction.set_rollback(True)
                        return JsonResponse(
                            {"result": 'bad', "mensaje": u"Error, el tamaño del archivo es mayor a 10 Mb."}, safe=False)
                    if not exte in ['pdf']:
                        transaction.set_rollback(True)
                        return JsonResponse({"result": 'bad', "mensaje": u"Error, solo archivo .pdf"}, safe=False)
                    if int(request.POST['tipo']) == 1:
                        arch = MemoActividadPosgrado.objects.get(pk=request.POST['id'])
                    elif int(request.POST['tipo']) == 2:
                        arch = InformeActividadJornada.objects.get(pk=request.POST['id'])
                    elif int(request.POST['tipo']) == 3:
                        arch = InformeTecnico.objects.get(pk=request.POST['id'])
                    elif int(request.POST['tipo']) == 4:
                        arch = ActaPago.objects.get(pk=request.POST['id'])
                    arch.archivofirmado = newfile
                    arch.save(request)
                    return JsonResponse({'result': 'ok', 'mensaje': 'Archivo guardado con éxito'}, safe=False)
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": 'bad', "mensaje": u"Error al procesar los datos."}, safe=False)

        elif action == 'traspasoconfirmado':
            try:
                responsable = int(request.POST['idd'])
                activo = int(request.POST['id'])
                acti = ActivoFijo.objects.filter(pk=activo, status=True)[0]
                solicitud = SolicitudActivos.objects.filter(pk=int(request.POST['idsolicitud']), status=True)[0]
                solicitud.estado = 2
                solicitud.save(request)
                data['activo'] = solicitud.activo
                data['solicitante'] = solicitud.solicitante
                qrname = 'Activo-' + str(activo)
                asunto = u"SOLICITUD DE TRASPASO DE ACTIVOS " + acti.descripcion
                send_html_mail(asunto, "emails/notificaractivo.html",
                               {'sistema': request.session['nombresistema'], 'Activo': acti,
                                'responsable': solicitud.responsableasignacion, 'activo': acti.descripcion,
                                'codigo': acti.codigogobierno},
                               ['dcevalloss1@unemi.edu.ec'], [],
                               cuenta=CUENTAS_CORREOS[16][1])
                log(u'Confirmó solicitud traspaso: %s' % solicitud, request, "add")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                pass

        elif action == 'confmantenimiento':
            try:
                solicitud = SolicitudConfirmacionMantenimiento.objects.get(pk=int(request.POST['id']))
                solicitud.estado = 2
                solicitud.save(request)
                log(u'Confirmó el mantenimiento realizado: %s' % solicitud, request, "add")
                historial = HistorialSolicitudConfirmacionMantenimiento(
                    solicitud=solicitud,
                    observacion=solicitud.observacion,
                    estado=solicitud.estado,
                    persona=persona
                )
                historial.save(request)
                log(u'Agrego registro de historial de confirmacion de mantenimiento %s' % (historial.__str__()),
                    request, "add")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)

        elif action == 'rechmantenimiento':
            try:
                solicitud = SolicitudConfirmacionMantenimiento.objects.get(pk=int(request.POST['id']))
                solicitud.estado = 3
                solicitud.save(request)
                log(u'Rechazó el mantenimiento realizado: %s' % solicitud, request, "add")
                historial = HistorialSolicitudConfirmacionMantenimiento(
                    solicitud=solicitud,
                    observacion=solicitud.observacion,
                    estado=solicitud.estado,
                    persona=persona
                )
                historial.save(request)
                log(u'Agrego registro de historial de confirmacion de mantenimiento %s' % (historial.__str__()),
                    request, "add")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)

        elif action == 'responsableconfirmatraspaso':
            try:
                if 'idsolicitud' in request.POST:
                    solicitud = SolicitudActivos.objects.filter(pk=int(request.POST['idsolicitud']), status=True)[0]
                    solicitud.estado = 4
                    solicitud.save(request)
                    log(u'Responsable confirmó solicitud traspaso: %s' % solicitud, request, "add")
                    return JsonResponse({"result": "ok"})
            except Exception as ex:
                pass

        elif action == 'traspasoeliminar':
            try:
                id = int(request.POST['id'])
                solicitud = SolicitudActivos.objects.get(pk=id, status=True)
                solicitud.estado = 3
                solicitud.save(request)
                log(u'Solicitud no confirmada: %s' % solicitud, request, "del")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                pass

        elif action == 'responsableeliminatraspaso':
            try:
                id = int(request.POST['id'])
                solicitud = SolicitudActivos.objects.get(pk=id, status=True)
                solicitud.estado = 5
                solicitud.save(request)
                log(u'Responsable rechaza solicitud traspaso: %s' % solicitud, request, "del")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                pass

        elif action == 'genera_acta_fam':
            try:
                familiares = persona.personadatosfamiliares_set.filter(status=True, aplicaproyeccion=True)
                if familiares:
                    familiares.update(actagenerada=True)

                    data['familiares'] =cantidad= familiares.count()
                    directory_p = os.path.join(MEDIA_ROOT, 'talento_humano')
                    try:
                        os.stat(directory_p)
                    except:
                        os.mkdir(directory_p)
                    directory = os.path.join(MEDIA_ROOT, 'talento_humano', 'actas_proyeccion')
                    nombre_archivo = generar_nombre(f'acta_proyeccion{request.user.username}', 'generado') + '.pdf'
                    try:
                        os.stat(directory)
                    except:
                        os.mkdir(directory)
                    data['persona'] = persona
                    data['fechaactual'] = datetime.now()
                    context = {'pagesize': 'A4 landscape', 'data': data}
                    valido = conviert_html_to_pdfsaveqr_generico(request,
                                                                 'th_hojavida/pdf/acta_cargafam.html',
                                                                 context,
                                                                 directory, nombre_archivo)
                    if not valido[0]:
                        raise NameError('Error al generar el acta')
                    url_archivo = f'talento_humano/actas_proyeccion/{nombre_archivo}'
                    archivo = Archivo(tipo_id=19, nombre = 'ACTA DE ACEPTACIÓN DE CARGAS FAMILIARES',
                                      fecha=datetime.now(),archivo=url_archivo, persona=persona)
                    archivo.save(request)

                    log(u'Generó acta por cargas familiares: %s cant: %s' % (persona,cantidad), request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('No registra familiares que apliquen proyección de gastos')

            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": f"{ex}"})

        elif action == 'addtransporte':
            try:
                with transaction.atomic():
                    form = PersonaTransporteForm(request.POST)
                    if form.is_valid():
                        if PersonaTransporte.objects.filter(persona=persona, transporte=form.cleaned_data['transporte'],
                                                            status=True).exists():
                            return JsonResponse({"result": "bad",
                                                 "mensaje": u"Tipo de transporte seleccionado ya existe en los registros."})
                        transportegasto = PersonaTransporte(persona=persona, transporte=form.cleaned_data['transporte'])
                        transportegasto.save(request)
                        log(u'Adicion transporte: %s' % transportegasto, request, "addtransporte")
                        return JsonResponse({"result": False, 'to': request.path}, safe=False)
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({"result": True, "mensaje": "Complete todos los campos vacios."},
                                            safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        elif action == 'edittransporte':
            try:
                with transaction.atomic():
                    transporte = PersonaTransporte.objects.get(pk=int(encrypt(request.POST['id'])))
                    form = PersonaTransporteForm(request.POST)
                    if form.is_valid():
                        if PersonaTransporte.objects.filter(persona=persona, transporte=form.cleaned_data['transporte'],
                                                            status=True).exclude(id=transporte.id).exists():
                            return JsonResponse({"result": "bad",
                                                 "mensaje": u"Tipo de transporte seleccionado ya existe en los registros."})
                        transporte.transporte = form.cleaned_data['transporte']
                        transporte.save(request)
                        log(u'Edicion de gasto por transporte: %s' % transporte, request, "edittransporte")
                        return JsonResponse({"result": False, 'to': request.path}, safe=False)
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({"result": True, "mensaje": "Complete todos los campos vacios."},
                                            safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        elif action == 'deltransporte':
            try:
                with transaction.atomic():
                    transporte = PersonaTransporte.objects.get(pk=int(encrypt(request.POST['id'])))
                    transporte.status = False
                    transporte.save(request)
                    log(u'Elimino registro de gasto por transporte: %s - %s - %s', request, "deltransporte")
                    res_json = {"error": False}
            except Exception as ex:
                res_json = {'error': True, "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)

        elif action == 'addalimentacion':
            try:
                with transaction.atomic():
                    form = PersonaAlimentacionUniversidadForm(request.POST)
                    if form.is_valid():
                        if PersonaAlimentacionUniversidad.objects.filter(persona=persona,
                                                                         lugar=form.cleaned_data['lugar'],
                                                                         status=True).exists():
                            return JsonResponse(
                                {"result": "bad", "mensaje": u"Lugar Seleccionado ya existe en los registros."})
                        alimentacion = PersonaAlimentacionUniversidad(persona=persona,
                                                                      lugar=form.cleaned_data['lugar'],
                                                                      )
                        alimentacion.save(request)
                        log(u'Agrego gasto de alimentacion en la universidad: %s' % alimentacion, request,
                            "addalimentacion")
                        return JsonResponse({"result": False, 'to': request.path}, safe=False)
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({"result": True, "mensaje": "Complete todos los campos vacios."},
                                            safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        elif action == 'editalimentacion':
            try:
                with transaction.atomic():
                    alimentacion = PersonaAlimentacionUniversidad.objects.get(pk=int(encrypt(request.POST['id'])))
                    form = PersonaAlimentacionUniversidadForm(request.POST)
                    if form.is_valid():
                        if PersonaAlimentacionUniversidad.objects.filter(persona=persona,
                                                                         lugar=form.cleaned_data['lugar'],
                                                                         status=True).exclude(
                                id=alimentacion.id).exists():
                            return JsonResponse(
                                {"result": "bad", "mensaje": u"Lugar Seleccionado ya existe en los registros."})
                        alimentacion.lugar = form.cleaned_data['lugar']
                        alimentacion.save(request)
                        log(u'Edicion de gasto por alimentacion en la universidad: %s' % alimentacion, request,
                            "editalimentacion")
                        return JsonResponse({"result": False, 'to': request.path}, safe=False)
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({"result": True, "mensaje": "Complete todos los campos vacios."},
                                            safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        elif action == 'delalimentacion':
            try:
                with transaction.atomic():
                    alimentacion = PersonaAlimentacionUniversidad.objects.get(pk=int(encrypt(request.POST['id'])))
                    alimentacion.status = False
                    alimentacion.save(request)
                    log(u'Elimino registro de gasto por alimentacion: %s - %s - %s', request, "delalimentacion")
                    res_json = {"error": False}
            except Exception as ex:
                res_json = {'error': True, "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)

        elif action == 'addplantelefonico':
            try:
                with transaction.atomic():
                    form = PersonaPlanTelefonicoForm(request.POST)
                    if form.is_valid():
                        plantelefonico = PersonaPlanTelefonico(persona=persona,
                                                               operadora=form.cleaned_data['operadora'],
                                                               tieneplan=form.cleaned_data['tieneplan'],
                                                               descripcion=form.cleaned_data['descripcion'],
                                                               )
                        plantelefonico.save(request)
                        log(u'Agrego gasto de plan telefonico: %s' % plantelefonico, request, "addplantelefonico")
                        return JsonResponse({"result": False, 'to': request.path}, safe=False)
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({"result": True, "mensaje": "Complete todos los campos vacios."},
                                            safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        elif action == 'editplantelefonico':
            try:
                with transaction.atomic():
                    plantelefonico = PersonaPlanTelefonico.objects.get(pk=int(encrypt(request.POST['id'])))
                    form = PersonaPlanTelefonicoForm(request.POST)
                    if form.is_valid():
                        plantelefonico.operadora = form.cleaned_data['operadora']
                        plantelefonico.tieneplan = form.cleaned_data['tieneplan']
                        plantelefonico.descripcion = form.cleaned_data['descripcion']
                        plantelefonico.save(request)
                        log(u'Edicion de gasto de plan telefonico: %s' % plantelefonico, request, "editplantelefonico")
                        return JsonResponse({"result": False, 'to': request.path}, safe=False)
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({"result": True, "mensaje": "Complete todos los campos vacios."},
                                            safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        elif action == 'delplantelefonico':
            try:
                with transaction.atomic():
                    plantelefonico = PersonaPlanTelefonico.objects.get(pk=int(encrypt(request.POST['id'])))
                    plantelefonico.status = False
                    plantelefonico.save(request)
                    log(u'Elimino registro de gasto por plan telefonico: %s - %s - %s', request, "delplantelefonico")
                    res_json = {"error": False}
            except Exception as ex:
                res_json = {'error': True, "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)

        elif action == 'addcompraalimentos':
            try:
                with transaction.atomic():
                    form = PersonaCompraAlimentosForm(request.POST)
                    if form.is_valid():
                        if PersonaCompraAlimentos.objects.filter(persona=persona, lugar=form.cleaned_data['lugar'],
                                                                 status=True).exists():
                            transaction.set_rollback(True)
                            return JsonResponse(
                                {"result": True, "mensaje": "Lugar seleccionado ya se encuentra registrado."},
                                safe=False)
                        compraalimentos = PersonaCompraAlimentos(persona=persona, lugar=form.cleaned_data['lugar'])
                        compraalimentos.save(request)
                        log(u'Agrego Gasto compra de alimentos: %s' % compraalimentos, request, "addcompraalimentos")
                        return JsonResponse({"result": False, 'to': request.path}, safe=False)
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({"result": True, "mensaje": "Complete todos los campos vacios."},
                                            safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        elif action == 'editcompraalimentos':
            try:
                with transaction.atomic():
                    compraalimentos = PersonaCompraAlimentos.objects.get(pk=int(encrypt(request.POST['id'])))
                    form = PersonaCompraAlimentosForm(request.POST)
                    if form.is_valid():
                        if PersonaCompraAlimentos.objects.filter(persona=persona, lugar=form.cleaned_data['lugar'],
                                                                 status=True).exclude(id=compraalimentos.id).exists():
                            transaction.set_rollback(True)
                            return JsonResponse(
                                {"result": True, "mensaje": "Lugar seleccionado ya se encuentra registrado."},
                                safe=False)
                        compraalimentos.lugar = form.cleaned_data['lugar']
                        compraalimentos.save(request)
                        log(u'Edicion de gasto de plan telefonico: %s' % compraalimentos, request,
                            "editcompraalimentos")
                        return JsonResponse({"result": False, 'to': request.path}, safe=False)
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({"result": True, "mensaje": "Complete todos los campos vacios."},
                                            safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        elif action == 'delcompraalimentos':
            try:
                with transaction.atomic():
                    compraalimento = PersonaCompraAlimentos.objects.get(pk=int(encrypt(request.POST['id'])))
                    compraalimento.status = False
                    compraalimento.save(request)
                    log(u'Elimino registro de compra de alimentos : %s - %s - %s', request, "delcompraalimentos")
                    res_json = {"error": False}
            except Exception as ex:
                res_json = {'error': True, "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)

        elif action == 'addgastoestudio':
            try:
                with transaction.atomic():
                    form = PersonaGastoMensualForm(request.POST)
                    if form.is_valid():
                        if PersonaGastoMensual.objects.filter(persona=persona, tipogasto=form.cleaned_data['tipogasto'],
                                                              status=True).exists():
                            transaction.set_rollback(True)
                            return JsonResponse(
                                {"result": True, "mensaje": "Gasto seleccionado ya se encuentra registrado."},
                                safe=False)
                        gasto = PersonaGastoMensual(persona=persona,
                                                    tipogasto=form.cleaned_data['tipogasto'],
                                                    valor=form.cleaned_data['valor'])
                        gasto.save(request)
                        log(u'Agrego Gasto: %s' % gasto, request, "addgasto")
                        return JsonResponse({"result": False, 'to': request.path}, safe=False)
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({"result": True, "mensaje": "Complete todos los campos vacios."},
                                            safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        elif action == 'editgastoestudio':
            try:
                with transaction.atomic():
                    gasto = PersonaGastoMensual.objects.get(pk=int(encrypt(request.POST['id'])))
                    form = PersonaGastoMensualForm(request.POST)
                    if form.is_valid():
                        if PersonaGastoMensual.objects.filter(persona=persona, tipogasto=form.cleaned_data['tipogasto'],
                                                              status=True).exclude(id=gasto.id).exists():
                            transaction.set_rollback(True)
                            return JsonResponse(
                                {"result": True, "mensaje": "Gasto seleccionado ya se encuentra registrado."},
                                safe=False)
                        gasto.tipogasto = form.cleaned_data['tipogasto']
                        gasto.valor = form.cleaned_data['valor']
                        gasto.save(request)
                        log(u'Edicion de gasto: %s' % gasto, request, "editgasto")
                        return JsonResponse({"result": False, 'to': request.path}, safe=False)
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({"result": True, "mensaje": "Complete todos los campos vacios."},
                                            safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        elif action == 'delgastoestudio':
            try:
                with transaction.atomic():
                    gasto = PersonaGastoMensual.objects.get(pk=int(encrypt(request.POST['id'])))
                    gasto.status = False
                    gasto.save(request)
                    log(u'Elimino registro de gasto : %s - %s - %s', request, "delgasto")
                    res_json = {"error": False}
            except Exception as ex:
                res_json = {'error': True, "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)

        elif action == 'adddocumentofirmado':
            try:
                with transaction.atomic():
                    newfile = None
                    form = AccionPersonalDocumentoForm(request.POST, request.FILES)
                    if form.is_valid():
                        accionpersonal = PersonaAcciones.objects.get(pk=int(request.POST['id']))
                        if 'archivofirmado' in request.FILES:
                            newfile = request.FILES['archivofirmado']
                            extension = newfile._name.split('.')
                            tam = len(extension)
                            exte = extension[tam - 1]
                            if newfile.size > 4194304:
                                return JsonResponse(
                                    {"result": True, "mensaje": u"Error, el tamaño del archivo es mayor a 4 Mb."})
                            if exte in ['pdf', 'jpg', 'jpeg', 'png', 'jpeg', 'peg']:
                                newfile._name = generar_nombre("AccionPersonalFirmado_", newfile._name)
                            else:
                                return JsonResponse(
                                    {"result": True, "mensaje": u"Error, solo archivos .pdf,.jpg, .jpeg"})
                            # accionpersonal.archivofirmado = newfile
                        historialdocumento = HistoricoDocumentosPersonaAcciones(personaaccion=accionpersonal,
                                                                                archivofirmado=newfile)
                        historialdocumento.save(request)
                        accionpersonal.estadoarchivo = 1
                        accionpersonal.save(request)

                        log(u'Adiciono Documento firmado en Accion de personal: %s' % accionpersonal, request, "add")
                        return JsonResponse({"result": False, 'to': request.path}, safe=False)
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({"result": True, "mensaje": "Complete los datos requeridos."}, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        elif action == 'adddocumentofirmadovacaciones':
            try:
                with transaction.atomic():
                    newfile = None
                    form = AccionPersonalDocumentoForm(request.POST, request.FILES)
                    if form.is_valid():
                        accionpersonal = AccionPersonal.objects.get(pk=int(request.POST['id']))
                        if 'archivofirmado' in request.FILES:
                            newfile = request.FILES['archivofirmado']
                            extension = newfile._name.split('.')
                            tam = len(extension)
                            exte = extension[tam - 1]
                            if newfile.size > 4194304:
                                return JsonResponse(
                                    {"result": True, "mensaje": u"Error, el tamaño del archivo es mayor a 4 Mb."})
                            if exte in ['pdf', 'jpg', 'jpeg', 'png', 'jpeg', 'peg']:
                                newfile._name = generar_nombre("AccionPersonalFirmado_", newfile._name)
                            else:
                                return JsonResponse(
                                    {"result": True, "mensaje": u"Error, solo archivos .pdf,.jpg, .jpeg"})
                        else:
                            return JsonResponse({"result": True, "mensaje": u"Error, no subió ningun documento"})
                            # accionpersonal.archivofirmado = newfile
                        historialdocumento = HistoricoDocumentosPersonaAcciones(personaaccionvacacion=accionpersonal,
                                                                                archivofirmado=newfile)
                        historialdocumento.save(request)
                        accionpersonal.estadoarchivo = 1
                        accionpersonal.save(request)

                        log(u'Adiciono Documento firmado en Accion de personal: %s' % accionpersonal, request, "add")
                        return JsonResponse({"result": False}, safe=False)
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({"result": True, "mensaje": "Complete los datos requeridos."}, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        elif action == 'editdocumentofirmadovacaciones':
            try:
                with transaction.atomic():
                    newfile = None
                    form = AccionPersonalDocumentoForm(request.POST, request.FILES)
                    if form.is_valid():
                        accionpersonal = AccionPersonal.objects.get(pk=int(request.POST['id']))
                        if accionpersonal.estadoarchivo == 1:
                            if 'archivofirmado' in request.FILES:
                                newfile = request.FILES['archivofirmado']
                                extension = newfile._name.split('.')
                                tam = len(extension)
                                exte = extension[tam - 1]
                                if newfile.size > 4194304:
                                    return JsonResponse(
                                        {"result": True, "mensaje": u"Error, el tamaño del archivo es mayor a 4 Mb."})
                                if exte in ['pdf', 'jpg', 'jpeg', 'png', 'jpeg', 'peg']:
                                    newfile._name = generar_nombre("AccionPersonalFirmado_", newfile._name)
                                else:
                                    return JsonResponse(
                                        {"result": True, "mensaje": u"Error, solo archivos .pdf,.jpg, .jpeg"})
                                # accionpersonal.archivofirmado = newfile
                            else:
                                return JsonResponse({"result": True, "mensaje": u"Error, no subió ningun documento"})
                            historialdocumento1 = HistoricoDocumentosPersonaAcciones.objects.filter(
                                personaaccionvacacion=accionpersonal, status=True).order_by('-id').last()
                            if historialdocumento1:
                                historialdocumento1.status = False
                                historialdocumento1.save()
                            historialdocumento = HistoricoDocumentosPersonaAcciones(
                                personaaccionvacacion=accionpersonal, archivofirmado=newfile)
                            historialdocumento.save(request)
                            accionpersonal.estadoarchivo = 1
                            accionpersonal.save(request)

                            log(u'Adiciono Documento firmado en Accion de personal: %s' % accionpersonal, request,
                                "add")
                            return JsonResponse({"result": False, 'to': request.path}, safe=False)
                        else:
                            return JsonResponse(
                                {"result": True, "mensaje": "El documento ya fue validado por Talento Humano"},
                                safe=False)

                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({"result": True, "mensaje": "Complete los datos requeridos."}, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        if action == 'activardasactivarcuentabancariabeca':
            try:
                with transaction.atomic():
                    val = request.POST['val']
                    id = int(request.POST['id'])
                    instance = CuentaBancariaPersona.objects.get(pk=id)
                    instance.activapago = val == 'y'
                    instance.save(request)
                    cuentas = persona.cuentabancariapersona_set.filter(status=True).exclude(pk=instance.id)
                    cuentas.update(activapago=False)
                    log(u'Actualizó estado del modelo %s plan de carrera: %s' % (instance, instance.activapago),
                        request, "edit")
                    res_json = {"error": False}
            except Exception as ex:
                transaction.set_rollback(True)
                res_json = {'error': True, "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)
        if action == 'importarActividades':
            try:
                with transaction.atomic():
                    if not 'archivo_excel' in request.FILES:
                        raise NameError('Carge un archivo excel para ejecutar la acción.')
                    excel = request.FILES['archivo_excel']
                    wb = openpyxl.load_workbook(excel)
                    worksheet = wb.worksheets[0]
                    count = 0
                    counter = 0
                    linea = 1
                    for row in worksheet.iter_rows():
                        currentValues = [str(cell.value) for cell in row]
                        if linea >= 2:
                            if currentValues[0].strip() in ('None', ''):
                                raise NameError(
                                    'REVISAR LINEA [{}],  FORMATO FILAS SIN DESCRIPCION DEL REQUERIMEINTO, VERIFIQUE QUE EL ARCHIVO NO TENGA COLUMNAS VACIAS AL FINAL.'.format(
                                        linea))
                            if currentValues[1].strip() in ('None', ''):
                                raise NameError(
                                    'REVISAR LINEA [{}],  FORMATO FILAS SIN FECHA DE INICIO, VERIFIQUE QUE EL ARCHIVO NO TENGA COLUMNAS VACIAS AL FINAL.'.format(
                                        linea))
                            if currentValues[2].strip() in ('None', ''):
                                raise NameError(
                                    'REVISAR LINEA [{}],  FORMATO FILAS SIN FECHA DE INICIO, VERIFIQUE QUE EL ARCHIVO NO TENGA COLUMNAS VACIAS AL FINAL.'.format(
                                        linea))
                            if currentValues[3].strip() in ('None', ''):
                                raise NameError(
                                    'REVISAR LINEA [{}],  FORMATO FILAS SIN NOMBRE DEL SISTEMA, VERIFIQUE QUE EL ARCHIVO NO TENGA COLUMNAS VACIAS AL FINAL.'.format(
                                        linea))
                            if currentValues[4].strip() in ('None', ''):
                                raise NameError(
                                    'REVISAR LINEA [{}],  FORMATO FILAS SIN TIPO DE ACTIVIDAD, VERIFIQUE QUE EL ARCHIVO NO TENGA COLUMNAS VACIAS AL FINAL.'.format(
                                        linea))
                            if currentValues[5].strip() in ('None', ''):
                                raise NameError(
                                    'REVISAR LINEA [{}],  FORMATO FILAS SIN DEPARTAMENTO, VERIFIQUE QUE EL ARCHIVO NO TENGA COLUMNAS VACIAS AL FINAL.'.format(
                                        linea))
                            requerimiento = currentValues[0].replace('\t', '').rstrip() if currentValues[
                                                                                               0] != 'None' else None
                            fechainicio = currentValues[1].replace('\t', '').rstrip() if currentValues[
                                                                                             1] != 'None' else None
                            fechafin = currentValues[2].replace('\t', '').rstrip() if currentValues[
                                                                                          2] != 'None' else None
                            sistema = currentValues[3].replace('\t', '').rstrip() if currentValues[
                                                                                         3] != 'None' else None
                            tipo = currentValues[4].replace('\t', '').rstrip() if currentValues[4] != 'None' else None
                            departamento = currentValues[5].replace('\t', '').rstrip() if currentValues[
                                                                                              5] != 'None' else None
                            tiposistema, tipoactividad = None, None
                            for s in TIPO_SISTEMA:
                                if s[1] == sistema:
                                    tiposistema = s[0]
                            for t in TIPO_ACTIVIDAD_BITACORA:
                                if t[1] == tipo:
                                    tipoactividad = t[0]
                            if departamento:
                                d = Departamento.objects.filter(nombre__icontains=departamento,
                                                                integrantes__isnull=False, status=True).distinct()
                            iddepartamento = d[0].id if d else 170
                            fechainicio = convertir_fecha_invertida(fechainicio.split(' ')[0])
                            fechafin = convertir_fecha_invertida(fechafin.split(' ')[0])
                            persona = Persona.objects.get(pk=request.POST['id'])

                            bitacora = BitacoraActividadDiaria(persona=persona,
                                                               fecha=fechainicio,
                                                               fechafin=fechafin,
                                                               departamento=persona.mi_departamento(),
                                                               descripcion=requerimiento,
                                                               # link=form.cleaned_data['link'],
                                                               actividades_id=17,
                                                               tiposistema=tiposistema,
                                                               # archivo=newfile,
                                                               departamento_requiriente_id=iddepartamento,
                                                               tipoactividad=tipoactividad
                                                               )
                            bitacora.save(request)
                            log(u'Adicionó bitacora de actividades diarias %s' % (bitacora), request, "add")
                            count += 1

                        linea += 1
                    messages.success(request, 'Importación de actividades realizada correctamente.')
                    return JsonResponse(
                        {"result": False, "mensaje": u"Importaciòn de actividades realizada correctamente."},
                        safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                print(ex)
                print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
                return JsonResponse({"result": "True", "mensaje": '%s' % (ex)})

        elif action == 'addhistoriallaboralaspirante':
            try:
                f = PersonaAportacionHistorialLaboralAspiranteForm(request.POST, request.FILES)
                if f.is_valid():
                    newfile = None
                    if 'archivo_resumen' in request.FILES:
                        newfile = request.FILES['archivo_resumen']
                        if newfile:
                            newfilesd = newfile._name
                            ext = newfilesd[newfilesd.rfind("."):]
                            if not ext == '.pdf':
                                return JsonResponse({"result": "bad", "mensaje": u"Error, solo archivos .pdf."})
                            if newfile.size > 20480000:
                                return JsonResponse({"result": "bad", "mensaje": u"Error, archivo mayor a 20 Mb."})
                            if newfile:
                                newfile._name = generar_nombre("archivo_resumen", newfile._name)

                    newfile2 = None
                    if 'archivo_detalle' in request.FILES:
                        newfile2 = request.FILES['archivo_detalle']
                        if newfile2:
                            newfilesd2 = newfile2._name
                            ext2 = newfilesd2[newfilesd2.rfind("."):]
                            if not ext2 == '.pdf':
                                return JsonResponse({"result": "bad", "mensaje": u"Error, solo archivos .pdf."})
                            if newfile2.size > 20480000:
                                return JsonResponse({"result": "bad", "mensaje": u"Error, archivo mayor a 20 Mb."})
                            if newfile2:
                                newfile2._name = generar_nombre("archivo_detalle", newfile2._name)

                    registro = PersonaAportacionHistorialLaboral(persona=persona)
                    if newfile:
                        registro.archivo_resumen = newfile
                    if newfile2:
                        registro.archivo_detalle = newfile2
                    registro.save(request)
                    log(u'Adicionó Historial Laboral de Aportación: %s' % registro, request, "add")
                    return JsonResponse({"result": False}, safe=False)
                else:
                    return JsonResponse({'error': True, "form": [{k: v[0]} for k, v in f.errors.items()],
                                         "message": "Error en el formulario"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'edithistoriallaboralaspirante':
            try:
                f = PersonaAportacionHistorialLaboralAspiranteForm(request.POST, request.FILES)
                if f.is_valid():
                    newfile = None
                    if 'archivo_resumen' in request.FILES:
                        newfile = request.FILES['archivo_resumen']
                        if newfile:
                            newfilesd = newfile._name
                            ext = newfilesd[newfilesd.rfind("."):]
                            if not ext == '.pdf':
                                return JsonResponse({"result": "bad", "mensaje": u"Error, solo archivos .pdf."})
                            if newfile.size > 20480000:
                                return JsonResponse({"result": "bad", "mensaje": u"Error, archivo mayor a 20 Mb."})
                            if newfile:
                                newfile._name = generar_nombre("archivo_resumen", newfile._name)

                    newfile2 = None
                    if 'archivo_detalle' in request.FILES:
                        newfile2 = request.FILES['archivo_detalle']
                        if newfile2:
                            newfilesd2 = newfile2._name
                            ext2 = newfilesd2[newfilesd2.rfind("."):]
                            if not ext2 == '.pdf':
                                return JsonResponse({"result": "bad", "mensaje": u"Error, solo archivos .pdf."})
                            if newfile2.size > 20480000:
                                return JsonResponse({"result": "bad", "mensaje": u"Error, archivo mayor a 20 Mb."})
                            if newfile2:
                                newfile2._name = generar_nombre("archivo_detalle", newfile2._name)

                    registro = PersonaAportacionHistorialLaboral.objects.get(pk=int(request.POST['id']))
                    if newfile:
                        registro.archivo_resumen = newfile
                    if newfile2:
                        registro.archivo_detalle = newfile2
                    registro.save(request)
                    log(u'Adicionó Historial Laboral de Aportación: %s' % registro, request, "add")
                    return JsonResponse({"result": False}, safe=False)
                else:
                    return JsonResponse({'error': True, "form": [{k: v[0]} for k, v in f.errors.items()],
                                         "message": "Error en el formulario"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        elif action == 'subirproyecciongastos':
            try:
                file=request.FILES['archivo']
                gasto=GastosPersonales.objects.get(id=request.POST['id'])
                name = file._name
                ext = name[name.rfind("."):]
                if not ext == '.pdf':
                    raise NameError('Solo se permite formato pdf')
                file._name = generar_nombre(f"proyeccion_{request.user.username}", "firmado")+'.pdf'
                gasto.archivo=file
                gasto.save(request)
                log(u'Guardo archivo firmado: {}'.format(gasto), request, "add")
                return JsonResponse({"result": False, "mensaje": "Guardado con exito", }, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": u"Error al guardar. %s" % ex.__str__()}, safe=False)

        # ACCIÓN DE PERSONAL
        elif action == 'firmadocvacaciones':
            try:
                # Parametros
                txtFirmas = json.loads(request.POST['txtFirmas'])
                if not txtFirmas:
                    raise NameError("Debe seleccionar ubicación de la firma")
                x = txtFirmas[-1]
                ida_personal = int(encrypt(request.POST['id_objeto']))
                accionpersonal = AccionPersonal.objects.get(pk=ida_personal)
                firma = request.FILES["firma"]
                passfirma = request.POST['palabraclave']
                url_archivo = (SITE_STORAGE + request.POST["url_archivo"]).replace('\\', '/')
                _name = generar_nombre(f'permiso_{request.user.username}_{ida_personal}_', 'firmada')
                folder = os.path.join(SITE_STORAGE, 'media', 'accionpersonalfirmados', '')
                # Firmar y guardar archivo en folder definido.
                # firma = firmar(request, passfirma, firma, pdf, x["numPage"],  x["x"], x["y"],x["width"], x["height"])
                firma = firmararchivogenerado(request, passfirma, firma, url_archivo, folder, _name, x["numPage"],
                                              x["x"], x["y"], x["width"], x["height"])
                if firma != True:
                    raise NameError(firma)
                log(u'Firmo Documento: {}'.format(_name), request, "add")

                folder_save = os.path.join('accionpersonalfirmados', '').replace('\\', '/')
                url_file_generado = f'{folder_save}{_name}.pdf'
                historialdocumento = HistoricoDocumentosPersonaAcciones(personaaccionvacacion=accionpersonal)
                historialdocumento.save(request)
                historialdocumento.archivofirmado = url_file_generado
                historialdocumento.save(request)
                accionpersonal.estadoarchivo = 1
                accionpersonal.save(request)
                log(u'Guardo archivo firmado: {}'.format(accionpersonal), request, "add")
                return JsonResponse({"result": False, "mensaje": "Guardado con exito", }, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": u"Error al guardar. %s" % ex.__str__()}, safe=False)

        elif action == 'firmaraccionpersonal':
            try:
                certificado = request.FILES["firma"]
                contrasenaCertificado = request.POST['palabraclave']
                razon = request.POST['razon'] if 'razon' in request.POST else ''
                extension_certificado = os.path.splitext(certificado.name)[1][1:]
                bytes_certificado = certificado.read()
                accionpersonal = AccionPersonal.objects.get(id=encrypt_id(request.POST['id']))
                firmas = []
                archivo_ = accionpersonal.archivo
                titulo = 'FUNCIONARIO O FUNCIONARIA'
                # Nota: El punto al final es esencial para el funcionamiento de la firma masiva, es el diferenciador ubicado en el html del certificado.
                palabras = f'{persona} {titulo}'.strip()
                x, y, numPage = obtener_posicion_x_y_saltolinea(archivo_.url, palabras, True, True)
                if x and y:
                    y = y - 10
                    firmas.append({'x': x, 'y': y, 'numPage': numPage})
                for membrete in firmas:
                    datau = JavaFirmaEc(
                        archivo_a_firmar=archivo_, archivo_certificado=bytes_certificado, extension_certificado=extension_certificado,
                        password_certificado=contrasenaCertificado,
                        page=int(membrete["numPage"]), reason=razon, lx=membrete["x"], ly=membrete["y"]
                    ).sign_and_get_content_bytes()
                    archivo_ = io.BytesIO()
                    archivo_.write(datau)
                    archivo_.seek(0)

                _name = f"accionpersonalfirmados_{len(accionpersonal.historial_documentos())}_{accionpersonal.id}_{accionpersonal.persona.usuario}"
                file_obj = DjangoFile(archivo_, name=f"{_name}.pdf")

                # accionpersonal.archivo = file_obj
                accionpersonal.estadoarchivo = 1
                accionpersonal.save(request)
                log(f'Firmo Documento de accion de personal: {accionpersonal} | {archivo_}', request, "add")

                historialdocumento = HistoricoDocumentosPersonaAcciones(personaaccionvacacion=accionpersonal,
                                                                        archivofirmado=file_obj)
                historialdocumento.save(request)

                log(u'Edito estado de accion de personal:  {}'.format(accionpersonal), request, "edit")
                return JsonResponse({'result': False, 'mensaje': 'Guardado correctamente'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, "mensaje": 'Error: {}'.format(ex)}, safe=False)

        # DATOS FAMILIARES
        elif action == 'addfamiliar':
            try:
                persona = request.session['persona']
                f = FamiliarForm(request.POST)
                if f.is_valid():
                    edit_d = eval(request.POST.get('edit_d', ''))
                    cedula = f.cleaned_data['identificacion'].strip()
                    if persona.personadatosfamiliares_set.filter(identificacion=f.cleaned_data['identificacion'], status=True).exists():
                        raise NameError('El familiar se encuentra registrado.')
                    nombres = f"{f.cleaned_data['apellido1']} {f.cleaned_data['apellido2']} {f.cleaned_data['nombre']}"
                    familiar = PersonaDatosFamiliares(persona=persona,
                                                      identificacion=cedula,
                                                      nombre=nombres,
                                                      fallecido=f.cleaned_data['fallecido'],
                                                      nacimiento=f.cleaned_data['nacimiento'],
                                                      parentesco=f.cleaned_data['parentesco'],
                                                      tienediscapacidad=f.cleaned_data['tienediscapacidad'],
                                                      telefono=f.cleaned_data['telefono'],
                                                      telefono_conv=f.cleaned_data['telefono_conv'],
                                                      niveltitulacion=f.cleaned_data['niveltitulacion'],
                                                      ingresomensual=f.cleaned_data['ingresomensual'],
                                                      formatrabajo=f.cleaned_data['formatrabajo'],
                                                      trabajo=f.cleaned_data['trabajo'],
                                                      convive=f.cleaned_data['convive'],
                                                      sustentohogar=f.cleaned_data['sustentohogar'],
                                                      # rangoedad=f.cleaned_data['rangoedad'],
                                                      essustituto=f.cleaned_data['essustituto'],
                                                      autorizadoministerio=f.cleaned_data['autorizadoministerio'],
                                                      tipodiscapacidad=f.cleaned_data['tipodiscapacidad'],
                                                      porcientodiscapacidad=f.cleaned_data['porcientodiscapacidad'],
                                                      carnetdiscapacidad=f.cleaned_data['carnetdiscapacidad'],
                                                      institucionvalida=f.cleaned_data['institucionvalida'],
                                                      tipoinstitucionlaboral=f.cleaned_data['tipoinstitucionlaboral'],
                                                      negocio=f.cleaned_data['negocio'],
                                                      esservidorpublico=f.cleaned_data['esservidorpublico'],
                                                      bajocustodia=f.cleaned_data['bajocustodia'],
                                                      centrocuidado=f.cleaned_data['centrocuidado'] if f.cleaned_data['centrocuidado'] else 0,
                                                      centrocuidadodesc=f.cleaned_data['centrocuidadodesc'],
                                                      tienenegocio=f.cleaned_data['tienenegocio'])
                    familiar.save(request)
                    if 'cedulaidentidad' in request.FILES:
                        newfile = request.FILES['cedulaidentidad']
                        newfile._name = generar_nombre("cedulaidentidad_", newfile._name)
                        familiar.cedulaidentidad = newfile
                        familiar.save(request)
                    if 'ceduladiscapacidad' in request.FILES:
                        newfile = request.FILES['ceduladiscapacidad']
                        newfile._name = generar_nombre("ceduladiscapacidad_", newfile._name)
                        familiar.ceduladiscapacidad = newfile
                        familiar.save(request)
                    if 'archivoautorizado' in request.FILES:
                        newfile = request.FILES['archivoautorizado']
                        newfile._name = generar_nombre("archivoautorizado_", newfile._name)
                        familiar.archivoautorizado = newfile
                        familiar.save(request)
                    if 'cartaconsentimiento' in request.FILES:
                        newfile = request.FILES['cartaconsentimiento']
                        newfile._name = generar_nombre(f"cartaconsentimiento_{persona.usuario.username}", newfile._name)
                        familiar.cartaconsentimiento = newfile
                        familiar.save(request)
                    if 'archivocustodia' in request.FILES:
                        newfile = request.FILES['archivocustodia']
                        newfile._name = generar_nombre(f"archivocustodia_{persona.usuario.username}", newfile._name)
                        familiar.archivocustodia = newfile
                        familiar.save(request)

                    pers = Persona.objects.filter(Q(pasaporte=cedula) | Q(cedula=cedula) | Q(pasaporte=('VS' + cedula)) | Q(cedula=cedula[2:]),status=True).first()
                    if not pers:
                        pers = Persona(cedula=f.cleaned_data['identificacion'],
                                       nombres=f.cleaned_data['nombre'],
                                       apellido1=f.cleaned_data['apellido1'],
                                       apellido2=f.cleaned_data['apellido2'],
                                       nacimiento=f.cleaned_data['nacimiento'],
                                       telefono=f.cleaned_data['telefono'],
                                       sexo=f.cleaned_data['sexo'],
                                       telefono_conv=f.cleaned_data['telefono_conv'],
                                       )
                        pers.save(request)
                        log(u'Adiciono persona: %s' % persona, request, "add")
                    elif len(pers.mis_perfilesusuarios()) == 1 and pers.tiene_usuario_externo():
                        pers.cedula = f.cleaned_data['identificacion']
                        pers.nombres = f.cleaned_data['nombre']
                        pers.apellido1 = f.cleaned_data['apellido1']
                        pers.apellido2 = f.cleaned_data['apellido2']
                        pers.nacimiento = f.cleaned_data['nacimiento']
                        pers.telefono = f.cleaned_data['telefono']
                        pers.sexo = f.cleaned_data['sexo']
                        pers.telefono_conv = f.cleaned_data['telefono_conv']
                        pers.save(request)
                        log(u'Edito familiar de usuario: %s' % pers, request, "edit")
                    if not pers.tiene_perfil():
                        if not Externo.objects.filter(persona=pers, status=True):
                            externo = Externo(persona=pers)
                            externo.save(request)
                            log(u'Adiciono externo: %s' % pers, request, "add")
                        perfil = PerfilUsuario(persona=pers, externo=externo)
                        perfil.save(request)
                        log(u'Adiciono perfil de usuario: %s' % perfil, request, "add")

                    perfil_i=pers.mi_perfil()
                    if not edit_d and perfil_i.tienediscapacidad:
                        familiar.tienediscapacidad=perfil_i.tienediscapacidad
                        familiar.tipodiscapacidad=perfil_i.tipodiscapacidad
                        familiar.porcientodiscapacidad=perfil_i.porcientodiscapacidad
                        familiar.carnetdiscapacidad=perfil_i.carnetdiscapacidad
                        familiar.institucionvalida=perfil_i.institucionvalida
                        familiar.ceduladiscapacidad=perfil_i.archivo.name if perfil_i.archivo else ''
                        familiar.archivoautorizado=perfil_i.archivovaloracion.name if perfil_i.archivovaloracion else ''
                    elif f.cleaned_data['tienediscapacidad']:
                        perfil_i.tienediscapacidad = f.cleaned_data['tienediscapacidad']
                        perfil_i.tipodiscapacidad = f.cleaned_data['tipodiscapacidad']
                        perfil_i.porcientodiscapacidad = f.cleaned_data['porcientodiscapacidad'] if f.cleaned_data['porcientodiscapacidad'] else 0
                        perfil_i.carnetdiscapacidad = f.cleaned_data['carnetdiscapacidad']
                        perfil_i.institucionvalida = f.cleaned_data['institucionvalida']
                        if 'ceduladiscapacidad' in request.FILES:
                            newfile = request.FILES['ceduladiscapacidad']
                            newfile._name = generar_nombre("archivosdiscapacidad_", newfile._name)
                            perfil_i.archivo = newfile
                            perfil_i.estadoarchivodiscapacidad = 1
                        if 'archivoautorizado' in request.FILES:
                            newfile = request.FILES['archivoautorizado']
                            newfile._name = generar_nombre("archivovaloracionmedica_", newfile._name)
                            perfil_i.archivovaloracion = newfile
                        perfil_i.save(request)
                    if f.cleaned_data['parentesco'].id in [14,11] and not persona.apellido1 in [pers.apellido1,pers.apellido2]:
                        familiar.aprobado=False
                    if f.cleaned_data['parentesco'].id == 13 and not pers.personadatosfamiliares_set.filter(personafamiliar=persona, status=True).exists():
                        fam_ = PersonaDatosFamiliares(persona=pers,
                                                      personafamiliar=persona,
                                                      identificacion=persona.cedula,
                                                      nombre=persona.nombre_completo_inverso(),
                                                      nacimiento=persona.nacimiento,
                                                      parentesco=f.cleaned_data['parentesco'],
                                                      telefono=persona.telefono,
                                                      telefono_conv=persona.telefono_conv,
                                                      convive=f.cleaned_data['convive'])
                        fam_.save(request)
                        perfil_fam = persona.mi_perfil()
                        if perfil_fam.tienediscapacidad:
                            fam_.tienediscapacidad = perfil_fam.tienediscapacidad
                            fam_.tipodiscapacidad = perfil_fam.tipodiscapacidad
                            fam_.porcientodiscapacidad = perfil_fam.porcientodiscapacidad
                            fam_.carnetdiscapacidad = perfil_fam.carnetdiscapacidad
                            fam_.institucionvalida = perfil_fam.institucionvalida
                            fam_.ceduladiscapacidad = perfil_fam.archivo.name if perfil_i.archivo else ''
                            fam_.archivoautorizado = perfil_fam.archivovaloracion.name if perfil_i.archivovaloracion else ''
                        fam_.save(request)
                    familiar.personafamiliar=pers
                    familiar.save(request)
                    if familiar.parentesco.id in [11, 14] or familiar.bajocustodia:
                        per_extension = persona.personaextension_set.filter(status=True).first()
                        hijos = per_extension.hijos if per_extension.hijos else 0
                        per_extension.hijos = hijos+1
                        per_extension.save(request)
                    log(u'Adiciono familiar: %s' % familiar, request, "add")
                    return JsonResponse({'result': False, 'mensaje':'Guardado con exito'})
                else:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()],
                                         "mensaje": "Error en el formulario"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', 'mensaje': f'{ex}'})

        elif action == 'editfamiliar':
            try:
                persona = request.session['persona']
                f = FamiliarForm(request.POST)
                f.edit()
                if f.is_valid():
                    familiar = PersonaDatosFamiliares.objects.get(pk=encrypt_id(request.POST['id']))
                    edit_d=eval(request.POST.get('edit_d',''))
                    cedula = f.cleaned_data['identificacion'].strip()
                    if persona.personadatosfamiliares_set.filter(identificacion=cedula, status=True).exclude(id=familiar.id).exists():
                        raise NameError(u'El familiar se encuentra registrado.')
                    nombres = f"{f.cleaned_data['apellido1']} {f.cleaned_data['apellido2']} {f.cleaned_data['nombre']}"
                    pers = Persona.objects.filter(Q(pasaporte=cedula) | Q(cedula=cedula) | Q(pasaporte=('VS' + cedula)) | Q(cedula=cedula[2:]),status=True).first()
                    familiar.fallecido = f.cleaned_data['fallecido']
                    familiar.parentesco = f.cleaned_data['parentesco']
                    familiar.tienediscapacidad = f.cleaned_data['tienediscapacidad']
                    familiar.trabajo = f.cleaned_data['trabajo']
                    familiar.niveltitulacion = f.cleaned_data['niveltitulacion']
                    familiar.ingresomensual = f.cleaned_data['ingresomensual']
                    familiar.formatrabajo = f.cleaned_data['formatrabajo']
                    familiar.convive = f.cleaned_data['convive']
                    familiar.sustentohogar = f.cleaned_data['sustentohogar']
                    # familiar.rangoedad = f.cleaned_data['rangoedad']
                    familiar.tienenegocio = f.cleaned_data['tienenegocio']
                    familiar.esservidorpublico = f.cleaned_data['esservidorpublico']
                    familiar.bajocustodia = f.cleaned_data['bajocustodia']
                    familiar.centrocuidado = f.cleaned_data['centrocuidado'] if f.cleaned_data['centrocuidado'] else 0
                    familiar.centrocuidadodesc = f.cleaned_data['centrocuidadodesc']
                    familiar.negocio = ''
                    familiar.tipoinstitucionlaboral = f.cleaned_data['tipoinstitucionlaboral']
                    if familiar.tienenegocio:
                        familiar.negocio = f.cleaned_data['negocio']
                    familiar.save(request)
                    if 'cedulaidentidad' in request.FILES:
                        newfile = request.FILES['cedulaidentidad']
                        newfile._name = generar_nombre("cedulaidentidad_", newfile._name)
                        familiar.cedulaidentidad = newfile
                        familiar.save(request)
                    if 'ceduladiscapacidad' in request.FILES:
                        newfile = request.FILES['ceduladiscapacidad']
                        newfile._name = generar_nombre("ceduladiscapacidad_", newfile._name)
                        familiar.ceduladiscapacidad = newfile
                        familiar.save(request)
                    if 'cartaconsentimiento' in request.FILES:
                        newfile = request.FILES['cartaconsentimiento']
                        newfile._name = generar_nombre(f"cartaconsentimiento_{persona.usuario.username}", newfile._name)
                        familiar.cartaconsentimiento = newfile
                        familiar.save(request)
                    if 'archivocustodia' in request.FILES:
                        newfile = request.FILES['archivocustodia']
                        newfile._name = generar_nombre(f"archivocustodia_{persona.usuario.username}", newfile._name)
                        familiar.archivocustodia = newfile
                        familiar.save(request)

                    if f.cleaned_data['tienediscapacidad']:
                        if 'archivoautorizado' in request.FILES:
                            newfile = request.FILES['archivoautorizado']
                            newfile._name = generar_nombre("archivoautorizado_", newfile._name)
                            familiar.archivoautorizado = newfile
                            familiar.save(request)
                    else:
                        familiar.archivoautorizado = None
                        familiar.save(request)
                    if not pers:
                        pers = Persona(cedula=f.cleaned_data['identificacion'],
                                       nombres=f.cleaned_data['nombre'],
                                       apellido1=f.cleaned_data['apellido1'],
                                       apellido2=f.cleaned_data['apellido2'],
                                       nacimiento=f.cleaned_data['nacimiento'],
                                       telefono=f.cleaned_data['telefono'],
                                       sexo=f.cleaned_data['sexo'],
                                       telefono_conv=f.cleaned_data['telefono_conv'],
                                       )
                        pers.save(request)
                        log(u'Adiciono persona: %s' % persona, request, "add")
                    elif len(pers.mis_perfilesusuarios()) == 1 and pers.tiene_usuario_externo():
                        pers.cedula = cedula
                        pers.nombres = f.cleaned_data['nombre']
                        pers.apellido1 = f.cleaned_data['apellido1']
                        pers.apellido2 = f.cleaned_data['apellido2']
                        pers.nacimiento = f.cleaned_data['nacimiento']
                        pers.telefono = f.cleaned_data['telefono']
                        pers.sexo = f.cleaned_data['sexo']
                        pers.telefono_conv = f.cleaned_data['telefono_conv']
                        pers.save(request)
                        log(u'Edito familiar de usuario: %s' % pers, request, "edit")
                    if not pers.tiene_perfil():
                        if not Externo.objects.filter(persona=pers, status=True):
                            externo = Externo(persona=pers)
                            externo.save(request)
                            log(u'Adiciono externo: %s' % pers, request, "add")
                        perfil = PerfilUsuario(persona=pers, externo=externo)
                        perfil.save(request)
                        log(u'Adiciono perfil de usuario: %s' % perfil, request, "add")
                    if len(pers.mis_perfilesusuarios()) == 1 and pers.tiene_usuario_externo():
                        familiar.identificacion = cedula
                        familiar.nombre = nombres
                        familiar.nacimiento = f.cleaned_data['nacimiento']
                        familiar.telefono = f.cleaned_data['telefono']
                        familiar.telefono_conv = f.cleaned_data['telefono_conv']
                    else:
                        familiar.identificacion = pers.cedula
                        familiar.nombre = pers.nombre_completo_inverso()
                        familiar.nacimiento = pers.nacimiento
                        familiar.telefono = pers.telefono
                        familiar.telefono_conv = pers.telefono_conv
                    perfil_i = pers.mi_perfil()
                    if edit_d:
                        if f.cleaned_data['tienediscapacidad']:
                            familiar.essustituto = f.cleaned_data['essustituto']
                            familiar.autorizadoministerio = f.cleaned_data['autorizadoministerio']
                            familiar.tipodiscapacidad = f.cleaned_data['tipodiscapacidad']
                            familiar.porcientodiscapacidad = f.cleaned_data['porcientodiscapacidad']
                            familiar.carnetdiscapacidad = f.cleaned_data['carnetdiscapacidad']
                            familiar.institucionvalida = f.cleaned_data['institucionvalida']
                            perfil_i.tienediscapacidad = f.cleaned_data['tienediscapacidad']
                            perfil_i.tipodiscapacidad = f.cleaned_data['tipodiscapacidad']
                            perfil_i.porcientodiscapacidad = f.cleaned_data['porcientodiscapacidad'] if f.cleaned_data['porcientodiscapacidad'] else 0
                            perfil_i.carnetdiscapacidad = f.cleaned_data['carnetdiscapacidad']
                            perfil_i.institucionvalida = f.cleaned_data['institucionvalida']
                            if 'ceduladiscapacidad' in request.FILES:
                                newfile = request.FILES['ceduladiscapacidad']
                                newfile._name = generar_nombre("archivosdiscapacidad_", newfile._name)
                                perfil_i.archivo = newfile
                                perfil_i.estadoarchivodiscapacidad = 1
                            if 'archivoautorizado' in request.FILES:
                                newfile = request.FILES['archivoautorizado']
                                newfile._name = generar_nombre("archivovaloracionmedica_", newfile._name)
                                perfil_i.archivovaloracion = newfile
                            perfil_i.save(request)
                        else:
                            familiar.essustituto = False
                            familiar.autorizadoministerio = False
                            familiar.tipodiscapacidad = None
                            familiar.porcientodiscapacidad = None
                            familiar.carnetdiscapacidad = ''
                            familiar.institucionvalida = None
                            perfil_i.tienediscapacidad = False
                            perfil_i.tipodiscapacidad = None
                            perfil_i.porcientodiscapacidad = None
                            perfil_i.carnetdiscapacidad = ''
                            perfil_i.institucionvalida = None
                            perfil_i.save(request)
                    elif perfil_i.tienediscapacidad:
                        familiar.essustituto = f.cleaned_data['essustituto']
                        familiar.autorizadoministerio = f.cleaned_data['autorizadoministerio']
                        familiar.tienediscapacidad = perfil_i.tienediscapacidad
                        familiar.tipodiscapacidad = perfil_i.tipodiscapacidad
                        familiar.porcientodiscapacidad = perfil_i.porcientodiscapacidad
                        familiar.carnetdiscapacidad = perfil_i.carnetdiscapacidad
                        familiar.institucionvalida = perfil_i.institucionvalida
                        familiar.ceduladiscapacidad = perfil_i.archivo.name if perfil_i.archivo else ''
                        familiar.archivoautorizado = perfil_i.archivovaloracion.name if perfil_i.archivovaloracion else ''
                        perfil_i.save(request)
                    if f.cleaned_data['parentesco'].id in [14, 11] and not persona.apellido1 in [pers.apellido1, pers.apellido2]:
                        familiar.aprobado = False
                    if f.cleaned_data['parentesco'].id == 13 and not pers.personadatosfamiliares_set.filter(personafamiliar=persona, status=True).exists():
                        fam_ = PersonaDatosFamiliares(persona=pers,
                                                      personafamiliar=persona,
                                                      identificacion=persona.cedula,
                                                      nombre=persona.nombre_completo_inverso(),
                                                      nacimiento=persona.nacimiento,
                                                      parentesco=f.cleaned_data['parentesco'],
                                                      telefono=persona.telefono,
                                                      telefono_conv=persona.telefono_conv,
                                                      convive=f.cleaned_data['convive'])
                        fam_.save(request)
                        perfil_fam = persona.mi_perfil()
                        if perfil_fam.tienediscapacidad:
                            fam_.tienediscapacidad = perfil_fam.tienediscapacidad
                            fam_.tipodiscapacidad = perfil_fam.tipodiscapacidad
                            fam_.porcientodiscapacidad = perfil_fam.porcientodiscapacidad
                            fam_.carnetdiscapacidad = perfil_fam.carnetdiscapacidad
                            fam_.institucionvalida = perfil_fam.institucionvalida
                            fam_.ceduladiscapacidad = perfil_fam.archivo.name if perfil_i.archivo else ''
                            fam_.archivoautorizado = perfil_fam.archivovaloracion.name if perfil_i.archivovaloracion else ''
                        fam_.save(request)
                    familiar.personafamiliar = pers
                    familiar.save(request)
                    log(u'Modifico familiar: %s' % persona, request, "edit")
                    return JsonResponse({'result': False, 'mensaje': 'Guardado con exito'})
                else:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()],
                                         "mensaje": "Error en el formulario"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', 'mensaje': str(ex)})

        elif action == 'subiractacargafamiliar':
            try:
                file=request.FILES['archivo']
                name = file._name
                ext = name[name.rfind("."):]
                if not ext == '.pdf':
                    raise NameError('Solo se permite formato pdf')
                file._name = generar_nombre(f"acta_proyeccion_{request.user.username}", "firmado")+'.pdf'
                archivo = Archivo(tipo_id=19, nombre='ACTA DE ACEPTACIÓN DE CARGAS FAMILIARES',
                                  fecha=datetime.now(), archivo=file, persona=persona, aprobado=True)
                archivo.save(request)
                log(u'Guardo archivo firmado: {}'.format(archivo), request, "add")
                return JsonResponse({"result": False, "mensaje": "Guardado con exito", }, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": u"Error al guardar. %s" % ex.__str__()}, safe=False)

        elif action == 'firmaactaproy':
            try:
                # Parametros
                txtFirmas = json.loads(request.POST['txtFirmas'])
                if not txtFirmas:
                    raise NameError("Debe seleccionar ubicación de la firma")
                x = txtFirmas[-1]
                ida_personal = int(encrypt(request.POST['id_objeto']))
                actaproyeccion = Archivo.objects.get(pk=ida_personal)
                firma = request.FILES["firma"]
                passfirma = request.POST['palabraclave']
                url_archivo = (SITE_STORAGE + request.POST["url_archivo"]).replace('\\', '/')
                _name = generar_nombre(f'acta_proyeccion{request.user.username}{ida_personal}', 'firmada')
                folder = os.path.join(SITE_STORAGE, 'media', 'talento_humano/actas_proyeccion', '')
                # Firmar y guardar archivo en folder definido.
                # firma = firmar(request, passfirma, firma, pdf, x["numPage"],  x["x"], x["y"],x["width"], x["height"])
                firma = firmararchivogenerado(request, passfirma, firma, url_archivo, folder, _name, x["numPage"],
                                              x["x"], x["y"], x["width"], x["height"])
                if firma != True:
                    raise NameError(firma)
                log(u'Firmo Documento: {}'.format(_name), request, "add")

                folder_save = os.path.join('talento_humano/actas_proyeccion', '').replace('\\', '/')
                url_file_generado = f'{folder_save}{_name}.pdf'

                actaproyeccion.aprobado = True
                actaproyeccion.archivo = url_file_generado
                actaproyeccion.save(request)
                log(u'Guardo archivo firmado: {}'.format(actaproyeccion), request, "add")
                return JsonResponse({"result": False, "mensaje": "Guardado con exito", }, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": u"Error al guardar. %s" % ex.__str__()}, safe=False)

        elif action == 'delfamiliar':
            try:
                persona = request.session['persona']
                familiar = PersonaDatosFamiliares.objects.get(pk=encrypt_id(request.POST['id']))
                familiar.status = False
                familiar.save(request)
                log(u'Elimino familiar: %s' % persona, request, "del")
                res_js = {'error': False}
            except Exception as ex:
                transaction.set_rollback(True)
                err = 'Error on line {}'.format(sys.exc_info()[-1].tb_lineno)
                msg_err = f"Ocurrio un error: {ex.__str__()}. {err}"
                res_js = {'error': True, 'mensaje': msg_err}
            return JsonResponse(res_js)

        #DATOS PERSONALES

        elif action == 'editdatospersonales':
            try:
                persona = request.session['persona']
                persona = Persona.objects.get(id=persona.id)
                f = DatosPersonalesForm(request.POST, request.FILES)
                if esestudiante:
                    f.es_estudiante()
                if not f.is_valid():
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()],
                                         "mensaje": "Error en el formulario"})
                if 'archivocedula' in request.FILES:
                    arch = request.FILES['archivocedula']
                    extension = arch._name.split('.')
                    tam = len(extension)
                    exte = extension[tam - 1]
                    if arch.size > 4194304:
                        raise NameError("Error, el tamaño del archivo de cédula es mayor a 4 Mb.")
                    if not exte.lower() == 'pdf':
                        raise NameError("Solo se permiten archivos .pdf")
                if 'archivoraza' in request.FILES:
                    arch_r = request.FILES['archivoraza']
                    extension = arch_r._name.split('.')
                    tam = len(extension)
                    exte = extension[tam - 1]
                    if arch_r.size > 4194304:
                        raise NameError("Error, el tamaño del archivo de etnia es mayor a 4 Mb.")
                    if not exte.lower() in ['pdf']:
                        raise NameError("Solo se permiten archivos .pdf")
                persona.pasaporte = f.cleaned_data['pasaporte']
                persona.anioresidencia = f.cleaned_data['anioresidencia']
                persona.nacimiento = f.cleaned_data['nacimiento']
                if not esestudiante:
                    persona.telefonoextension = f.cleaned_data['extension']
                persona.sexo = f.cleaned_data['sexo']
                # persona.agendacitas = f.cleaned_data['agendacitas']
                persona.lgtbi = f.cleaned_data['lgtbi']
                persona.email = f.cleaned_data['email']
                persona.libretamilitar = f.cleaned_data['libretamilitar']
                persona.eszurdo = f.cleaned_data['eszurdo']
                persona.save(request)
                personaextension = persona.datos_extension()
                personaextension.estadocivil = f.cleaned_data['estadocivil']
                personaextension.save(request)

                if 'archivocedula' in request.FILES:
                    newfile = request.FILES['archivocedula']
                    newfile._name = generar_nombre("cedula", newfile._name)

                    documento = persona.documentos_personales()
                    if documento is None:
                        documento = PersonaDocumentoPersonal(persona=persona,
                                                             cedula=newfile,
                                                             estadocedula=1
                                                             )
                    else:
                        documento.cedula = newfile
                        documento.estadocedula = 1

                    documento.save(request)

                if 'papeleta' in request.FILES:
                    newfile = request.FILES['papeleta']
                    newfile._name = generar_nombre("papeleta", newfile._name)

                    documento = persona.documentos_personales()
                    if documento is None:
                        documento = PersonaDocumentoPersonal(persona=persona,
                                                             papeleta=newfile,
                                                             estadopapeleta=1
                                                             )
                    else:
                        documento.papeleta = newfile
                        documento.estadopapeleta = 1

                    documento.save(request)

                if 'archivolibretamilitar' in request.FILES:
                    newfile = request.FILES['archivolibretamilitar']
                    newfile._name = generar_nombre("libretamilitar", newfile._name)

                    documento = persona.documentos_personales()
                    if documento is None:
                        documento = PersonaDocumentoPersonal(persona=persona,
                                                             libretamilitar=newfile,
                                                             estadolibretamilitar=1
                                                             )
                    else:
                        documento.libretamilitar = newfile
                        documento.estadolibretamilitar = 1

                    documento.save(request)

                perfil = persona.mi_perfil()
                perfil.raza = f.cleaned_data['raza']
                perfil.nacionalidadindigena = f.cleaned_data['nacionalidadindigena']
                if 'archivoraza' in request.FILES:
                    arch_r._name = generar_nombre("archivoraza", arch_r._name)
                    perfil.archivoraza = arch_r
                    perfil.estadoarchivoraza = 1
                else:
                    if perfil.raza.id != 1:
                        perfil.archivoraza = None
                        perfil.estadoarchivoraza = None
                perfil.save(request)
                request.session['persona'] = persona
                log(u'Modifico datos personales: %s' % persona, request, "edit")
                return JsonResponse({'result': False, 'mensaje': 'Guardado con exito'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': f'Error al guardar los datos: {ex}'})

        elif action == 'editdatosnacimiento':
            try:
                persona = request.session['persona']
                f = DatosNacimientoForm(request.POST)
                if not f.is_valid():
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()],
                                         "mensaje": "Error en el formulario"})
                persona.paisnacimiento = f.cleaned_data['paisnacimiento']
                persona.provincianacimiento = f.cleaned_data['provincianacimiento']
                persona.cantonnacimiento = f.cleaned_data['cantonnacimiento']
                persona.parroquianacimiento = f.cleaned_data['parroquianacimiento']
                persona.paisnacionalidad = f.cleaned_data['paisnacionalidad']
                persona.save(request)
                log(u'Modifico datos de nacimiento: %s' % persona, request, "edit")
                return JsonResponse({'result': False, 'mensaje': 'Guardado con exito'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': f'Error al guardar los datos {ex}'})

        elif action == 'editdatosdomicilio':
            try:
                if 'archivocroquis' in request.FILES:
                    newfile = request.FILES['archivocroquis']
                    if newfile.size > 2194304:
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"Error, Tamaño de archivo Maximo permitido es de 2Mb"})

                if 'archivoplanillaluz' in request.FILES:
                    newfile = request.FILES['archivoplanillaluz']
                    if newfile.size > 2194304:
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"Error, Tamaño de archivo Maximo permitido es de 2Mb"})

                if 'serviciosbasico' in request.FILES:
                    newfile2 = request.FILES['serviciosbasico']
                    if newfile2.size > 2194304:
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"Error, Tamaño de archivo Maximo permitido es de 2Mb"})

                persona = request.session['persona']
                documento_p=persona.documentos_personales()
                f = DatosDomicilioForm(request.POST)
                if 'pais' in request.POST and request.POST['pais'] and int(request.POST['pais']) == 1:
                    if 'provincia' in request.POST and not request.POST['provincia']:
                        raise NameError('Debe ingresa una provincia')
                    if 'canton' in request.POST and not request.POST['canton']:
                        raise NameError('Debe ingresa una canton')
                    if 'parroquia' in request.POST and not request.POST['parroquia']:
                        raise NameError('Debe ingresa una parroquia')
                if not f.is_valid():
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()],
                                         "mensaje": "Error en el formulario"})

                newfile = None
                persona.pais = f.cleaned_data['pais']
                persona.provincia = f.cleaned_data['provincia']
                persona.canton = f.cleaned_data['canton']
                persona.sector = f.cleaned_data['sector']
                persona.parroquia = f.cleaned_data['parroquia']
                persona.direccion = f.cleaned_data['direccion']
                persona.direccion2 = f.cleaned_data['direccion2']
                persona.ciudadela = f.cleaned_data['ciudadela']
                persona.num_direccion = f.cleaned_data['num_direccion']
                persona.telefono_conv = f.cleaned_data['telefono_conv']
                persona.telefono = f.cleaned_data['telefono']
                persona.tipocelular = f.cleaned_data['tipocelular']
                persona.referencia = f.cleaned_data['referencia']
                persona.zona = int(f.cleaned_data['zona'])
                persona.save(request)
                if 'archivocroquis' in request.FILES:
                    newfile = request.FILES['archivocroquis']
                    newfile._name = generar_nombre("croquis_", newfile._name)
                    persona.archivocroquis = newfile
                    persona.save(request)

                if 'archivoplanillaluz' in request.FILES:
                    newfile = request.FILES['archivoplanillaluz']
                    newfile._name = generar_nombre("planilla_luz_", newfile._name)
                    persona.archivoplanillaluz = newfile
                    persona.save(request)
                if 'serviciosbasico' in request.FILES:
                    newfile = request.FILES['serviciosbasico']
                    newfile._name = generar_nombre("serviciobasico_", newfile._name)
                    documento_p.serviciosbasico = newfile
                    documento_p.estadoserviciosbasico = 1
                    documento_p.save(request)

                log(u'Modifico datos de domicilio: %s' % persona, request, "edit")
                return JsonResponse({'result': False,'mensaje':'Guardado con éxito'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': f'Error al guardar los datos {ex}'})

        elif action == 'datosmigrante':
            try:
                if 'archivo' in request.FILES:
                    arch = request.FILES['archivo']
                    extension = arch._name.split('.')
                    tam = len(extension)
                    exte = extension[tam - 1]
                    if arch.size > 4194304:
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"Error, el tamaño del archivo es mayor a 4 Mb."})
                    if not exte.lower() in ['pdf']:
                        return JsonResponse({"result": "bad", "mensaje": u"Solo se permiten archivos .pdf"})

                form = MigranteForm(request.POST, request.FILES)
                if not form.is_valid():
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()],
                                         "mensaje": "Error en el formulario"})
                if not MigrantePersona.objects.filter(persona=persona).exists():
                    migrante = MigrantePersona(persona=persona,
                                               paisresidencia=form.cleaned_data['paisresidencia'],
                                               anioresidencia=form.cleaned_data['anioresidencia'],
                                               mesresidencia=form.cleaned_data['mesresidencia'],
                                               fecharetorno=form.cleaned_data['fecharetorno']
                                               )
                    migrante.save(request)

                    if 'archivo' in request.FILES:
                        arch._name = generar_nombre("archivomigrante", arch._name)
                        migrante.archivo = arch
                        migrante.estadoarchivo = 1
                        migrante.save(request)
                    log(u'Agregó datos de migrante retornado: %s - la persona: %s' % (migrante, persona), request,
                        "add")
                else:
                    migrante = MigrantePersona.objects.get(persona=persona)
                    migrante.paisresidencia = form.cleaned_data['paisresidencia']
                    migrante.anioresidencia = form.cleaned_data['anioresidencia']
                    migrante.mesresidencia = form.cleaned_data['mesresidencia']
                    migrante.fecharetorno = form.cleaned_data['fecharetorno']

                    if 'archivo' in request.FILES:
                        arch._name = generar_nombre("archivomigrante", arch._name)
                        migrante.archivo = arch
                        migrante.estadoarchivo = 1

                    migrante.save(request)
                    log(u'Editó datos de migrante retornado: %s - la persona: %s' % (migrante, persona), request,
                        "edit")

                return JsonResponse({"result": False})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": False, "mensaje": f"Error al guardar los datos:{ex}"})

        # DECLARACIONES
        # elif action == 'adddeclaracion':
        #     try:
        #         newfile = None
        #         if 'archivo' in request.FILES:
        #             newfile = request.FILES['archivo']
        #             if newfile.size > 4194304:
        #                 return JsonResponse(
        #                     {"result": "bad", "mensaje": u"Error, Tamaño de archivo Maximo permitido es de 4Mb"})
        #         persona = request.session['persona']
        #         f = DeclaracionBienAspiranteForm(request.POST, request.FILES)
        #         if f.is_valid():
        #             if 'archivo' not in request.FILES:
        #                 return JsonResponse({'result': 'bad', 'mensaje': u'Debe subir la declaración en formato pdf'})
        #             if DeclaracionBienes.objects.filter(status=True, persona=persona, fecha=f.cleaned_data['fecha']).exists():
        #                 return JsonResponse({'result': 'bad', 'mensaje': u'Ya tiene una declaración en la misma fecha'})
        #             newfile = request.FILES['archivo']
        #             newfile._name = generar_nombre("declaracion_", newfile._name)
        #             fechaproxima = None
        #             if int(f.cleaned_data['tipodeclaracion']) == 3:
        #                 fechaproxima = f.cleaned_data['fecha'] + timedelta(days=730)  # sumar dos años
        #
        #             declaracion = DeclaracionBienes(persona=persona,
        #                                             # numero=f.cleaned_data['numero'],
        #                                             # provincia=f.cleaned_data['provincia'],
        #                                             # canton=f.cleaned_data['canton'],
        #                                             # parroquia=f.cleaned_data['parroquia'],
        #                                             fecha=f.cleaned_data['fecha'],
        #                                             tipodeclaracion=f.cleaned_data['tipodeclaracion'],
        #                                             codigobarra=f.cleaned_data['codigobarra'],
        #                                             fechaproximoregistro=fechaproxima,
        #                                             archivo=newfile)
        #             declaracion.save(request)
        #             if newfile is not None:
        #                 if DistributivoPersona.objects.filter(status=True, persona=persona).exists():
        #                     lista = ['talento_humano@unemi.edu.ec', 'portizg@unemi.edu.ec']
        #                     asunto = "Ingresaron nueva declaración (archivo)"
        #                     send_html_mail(asunto, "emails/ingreso_archivos_tthh.html",
        #                                    {'asunto': asunto, 'd': declaracion.persona.nombre_completo_inverso(),
        #                                     'fecha': datetime.now().date(), 'escenario': 'declaración'}, lista, [],
        #                                    cuenta=CUENTAS_CORREOS[1][1])
        #                 log(u'Adiciono declaracion: %s' % persona, request, "add")
        #             return JsonResponse({'result': False, 'mensaje': 'Guardado con exito'})
        #         else:
        #             transaction.set_rollback(True)
        #             return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()],
        #                                      "mensaje": "Error en el formulario"})
        #     except Exception as ex:
        #         transaction.set_rollback(True)
        #         return JsonResponse({'result': True, 'mensaje': f'{ex}'})

        # CUENTA BANCARIA

        elif action == 'addcuentabancaria':
            try:
                f = CuentaBancariaPersonaForm(request.POST)
                if f.is_valid():
                    numero = f.cleaned_data['numero'].strip()
                    if persona.cuentabancariapersona_set.filter(numero=numero).exists():
                        return JsonResponse({'result': True,
                                             "mensaje": "La cuenta bancaria se encuentra registrada."})
                    cuentabancaria = CuentaBancariaPersona(persona=persona,
                                                           numero=f.cleaned_data['numero'],
                                                           estadorevision=1,
                                                           banco=f.cleaned_data['banco'],
                                                           tipocuentabanco=f.cleaned_data['tipocuentabanco'], )
                    cuentabancaria.save(request)
                    if 'archivo' in request.FILES:
                        newfile = request.FILES['archivo']
                        if newfile:
                            newfile._name = generar_nombre("cuentabancaria_", newfile._name)
                            cuentabancaria.archivo = newfile
                            cuentabancaria.save(request)
                        if DistributivoPersona.objects.filter(status=True,persona=persona ):
                            lista = ['talento_humano@unemi.edu.ec', 'portizg@unemi.edu.ec']
                            asunto = "Ingresaron nueva cuenta bancaria (archivo)"
                            send_html_mail(asunto, "emails/ingreso_archivos_tthh.html",
                                           {'asunto': asunto, 'd': cuentabancaria.persona.nombre_completo_inverso(),
                                            'fecha': datetime.now().date(), 'escenario': 'cuenta bancaria'}, lista, [],
                                           cuenta=CUENTAS_CORREOS[1][1])
                    log(u'Adiciono cuenta bancaria: %s' % persona, request, "add")
                    return JsonResponse({'result': False, 'mensaje': 'Guardado con exito'})
                else:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()],
                                         "mensaje": "Error en el formulario"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', 'mensaje': f'{ex}'})

        elif action == 'editcuentabancaria':
            try:
                persona = request.session['persona']
                f = CuentaBancariaPersonaForm(request.POST)
                if f.is_valid():
                    cuentabancaria = CuentaBancariaPersona.objects.get(pk=int(encrypt(request.POST['id'])))
                    if persona.cuentabancariapersona_set.filter(numero=f.cleaned_data['numero'].strip()).exclude(
                            id=cuentabancaria.id).exists():
                        return JsonResponse({'result': True,
                                             "mensaje": "La cuenta bancaria se encuentra registrada."})

                    if cuentabancaria.verificado:
                        return JsonResponse({'result': True,
                                             "mensaje": "mensaje': u'No puede modificar la cuenta bancaria."})
                    cuentabancaria.numero = f.cleaned_data['numero']
                    cuentabancaria.banco = f.cleaned_data['banco']
                    cuentabancaria.tipocuentabanco = f.cleaned_data['tipocuentabanco']
                    cuentabancaria.save(request)
                    if 'archivo' in request.FILES:
                        newfile = request.FILES['archivo']
                        if newfile:
                            newfile._name = generar_nombre("cuentabancaria_", newfile._name)
                            cuentabancaria.archivo = newfile
                            cuentabancaria.save(request)
                    log(u'Modifico cuenta bancaria: %s' % persona, request, "edit")
                    return JsonResponse({'result': False, 'mensaje': 'Guardado con exito'})
                else:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()],
                                         "mensaje": "Error en el formulario"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', 'mensaje': f'{ex}'})

        elif action == 'delcuentabancaria':
            try:
                cuentabancaria = CuentaBancariaPersona.objects.get(pk=int(encrypt(request.POST['id'])))
                if cuentabancaria.verificado:
                    res_js = {'error': True, 'mensaje': 'No puede eliminar la cuenta bancaria.'}
                    return JsonResponse(res_js)
                if cuentabancaria.activapago:
                    res_js = {'error': True, 'mensaje': 'No puede eliminar la cuenta bancaria se encuentra asignada al proceso de beca.'}
                    return JsonResponse(res_js)
                cuentabancaria.status = False
                cuentabancaria.save(request)
                log(u'Elimino cuenta bancaria: %s' % cuentabancaria, request, "del")
                res_js = {'error': False}
            except Exception as ex:
                transaction.set_rollback(True)
                err = 'Error on line {}'.format(sys.exc_info()[-1].tb_lineno)
                msg_err = f"Ocurrio un error: {ex.__str__()}. {err}"
                res_js = {'error': True, 'mensaje': msg_err}
            return JsonResponse(res_js)
        
        # REFERENCIAS
        
        elif action == 'addreferencia':
            try:
                form = ReferenciaPersonaForm(request.POST)
                if form.is_valid():
                    referenciapersona = ReferenciaPersona(persona=persona,
                                                          nombres=form.cleaned_data['nombres'],
                                                          apellidos=form.cleaned_data['apellidos'],
                                                          email=form.cleaned_data['email'],
                                                          telefono=form.cleaned_data['telefono'],
                                                          institucion=form.cleaned_data['institucion'],
                                                          relacion=form.cleaned_data['relacion'],
                                                          cargo=form.cleaned_data['cargo'])
                    referenciapersona.save(request)
                    log(u'Adicionó un nuevo referencia personal a la hoja de vida: %s - la persona: %s' % (
                        referenciapersona, persona), request, "add")
                    return JsonResponse({'result': False, 'mensaje': 'Guardado con exito'})
                else:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()],
                                         "mensaje": "Error en el formulario"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', 'mensaje': f'{ex}'})

        elif action == 'editreferencia':
            try:
                form = ReferenciaPersonaForm(request.POST)
                if form.is_valid():
                    referenciapersona = ReferenciaPersona.objects.get(pk=encrypt(request.POST['id']))
                    referenciapersona.nombres = form.cleaned_data['nombres']
                    referenciapersona.apellidos = form.cleaned_data['apellidos']
                    referenciapersona.email = form.cleaned_data['email']
                    referenciapersona.telefono = form.cleaned_data['telefono']
                    referenciapersona.institucion = form.cleaned_data['institucion']
                    referenciapersona.relacion = form.cleaned_data['relacion']
                    referenciapersona.cargo = form.cleaned_data['cargo']
                    referenciapersona.save(request)
                    log(u'Editó una referencia personal a la hoja de vida: %s - la persona: %s' % (
                        referenciapersona, persona), request, "edit")
                    return JsonResponse({'result': False, 'mensaje': 'Guardado con exito'})
                else:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()],
                                         "mensaje": "Error en el formulario"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', 'mensaje': f'{ex}'})

        elif action == 'delreferencia':
            try:
                referenciapersona = ReferenciaPersona.objects.get(pk=encrypt(request.POST['id']))
                log(u'Eliminó un referencia personal a la hoja de vida: %s - la persona: %s' % (
                    referenciapersona, persona), request, "add")
                referenciapersona.status = False
                referenciapersona.save(request)
                res_js = {'error': False}
            except Exception as ex:
                transaction.set_rollback(True)
                err = 'Error on line {}'.format(sys.exc_info()[-1].tb_lineno)
                msg_err = f"Ocurrio un error: {ex.__str__()}. {err}"
                res_js = {'error': True, 'mensaje': msg_err}
            return JsonResponse(res_js)

        # DATOS MÉDICOS

        elif action == 'addembarazo':
            try:
                f = PersonaDetalleMaternidadForm(request.POST)
                if f.is_valid():
                    embarazo = PersonaDetalleMaternidad(persona=persona,
                                                        gestacion=f.cleaned_data['gestacion'],
                                                        semanasembarazo=f.cleaned_data['semanasembarazo'],
                                                        lactancia=f.cleaned_data['lactancia'],
                                                        fechaparto=f.cleaned_data['fechaparto'],
                                                        fechainicioembarazo=f.cleaned_data['fechainicioembarazo'],
                                                        status_gestacion=f.cleaned_data['gestacion'],
                                                        status_lactancia=f.cleaned_data['lactancia'])
                    embarazo.save(request)

                    log(u'Adiciono embarazo: %s' % embarazo, request, "add")
                    return JsonResponse({'result': False, 'mensaje': 'Guardado con exito'})
                else:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()],
                                         "mensaje": "Error en el formulario"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', 'mensaje': f'{ex}'})

        elif action == 'editembarazo':
            try:
                f = PersonaDetalleMaternidadForm(request.POST)
                embarazo = PersonaDetalleMaternidad.objects.get(pk=int(encrypt(request.POST['id'])))
                if f.is_valid():
                    embarazo.persona=persona
                    embarazo.gestacion=f.cleaned_data['gestacion']
                    embarazo.semanasembarazo=f.cleaned_data['semanasembarazo']
                    embarazo.lactancia=f.cleaned_data['lactancia']
                    embarazo.fechaparto=f.cleaned_data['fechaparto']
                    embarazo.fechainicioembarazo=f.cleaned_data['fechainicioembarazo']
                    embarazo.status_gestacion=f.cleaned_data['gestacion']
                    embarazo.status_lactancia=f.cleaned_data['lactancia']
                    embarazo.save(request)

                    log(u'editó embarazo: %s' % embarazo, request, "edit")
                    return JsonResponse({'result': False, 'mensaje': 'Guardado con exito'})
                else:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()],
                                         "mensaje": "Error en el formulario"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', 'mensaje': f'{ex}'})

        elif action == 'delembarazo':
            try:
                embarazo = PersonaDetalleMaternidad.objects.get(pk=encrypt_id(request.POST['id']))
                embarazo.status = False
                embarazo.save(request)
                log(u'Elimino embarazo: %s' % persona, request, "del")
                res_js = {'error': False}
            except Exception as ex:
                transaction.set_rollback(True)
                err = 'Error on line {}'.format(sys.exc_info()[-1].tb_lineno)
                msg_err = f"Ocurrio un error: {ex.__str__()}. {err}"
                res_js = {'error': True, 'mensaje': msg_err}
            return JsonResponse(res_js)

        elif action == 'editdatosmedicos':
            try:
                persona = request.session['persona']
                f = DatosMedicosForm(request.POST)
                if f.is_valid():
                    datosextension = persona.datos_extension()
                    examenfisico = persona.datos_examen_fisico()
                    datosextension.carnetiess = f.cleaned_data['carnetiess']
                    personaexamenfisico = \
                        PersonaExamenFisico.objects.filter(personafichamedica__personaextension=datosextension)[0]
                    personaexamenfisico.peso = f.cleaned_data['peso']
                    personaexamenfisico.talla = f.cleaned_data['talla']
                    personaexamenfisico.save(request)
                    persona.sangre = f.cleaned_data['sangre']
                    persona.save(request)
                    datosextension.save(request)

                    if 'archivotiposangre' in request.FILES:
                        newfile = request.FILES['archivotiposangre']
                        newfile._name = generar_nombre("tiposangre", newfile._name)

                        documento = persona.documentos_personales()
                        if documento is None:
                            documento = PersonaDocumentoPersonal(persona=persona,
                                                                 tiposangre=newfile,
                                                                 estadotiposangre=1
                                                                 )
                        else:
                            documento.tiposangre = newfile
                            documento.estadotiposangre = 1

                        documento.save(request)

                    log(u'Modifico datos de medicos basicos: %s' % persona, request, "edit")
                    return JsonResponse({'result': False, 'mensaje': 'Guardado con exito'})
                else:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()],
                                     "mensaje": "Error en el formulario"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', 'mensaje': f'{ex}'})

        elif action == 'editcontactoemergencia':
            try:
                f = ContactoEmergenciaForm(request.POST)
                if f.is_valid():
                    datosextension = persona.datos_extension()
                    datosextension.contactoemergencia=f.cleaned_data['contactoemergencia']
                    datosextension.parentescoemergencia=f.cleaned_data['parentescoemergencia']
                    datosextension.telefonoemergencia=f.cleaned_data['telefonoemergencia']
                    datosextension.correoemergencia=f.cleaned_data['correoemergencia']
                    datosextension.telefonoconvemergencia=f.cleaned_data['telefonoconvemergencia']
                    datosextension.save(request)
                    log(u'editó contacto de emergencia: %s' % datosextension, request, "edit")
                    return JsonResponse({'result': False, 'mensaje': 'Guardado con exito'})
                else:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()],
                                         "mensaje": "Error en el formulario"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', 'mensaje': f'{ex}'})

        elif action == 'addvacunacvd19':
            try:
                with transaction.atomic():
                    newfile = None
                    # vacunado = False
                    form = VacunacionCovidForm(request.POST, request.FILES)
                    if form.is_valid():
                        vacunacion = VacunaCovid(persona=persona, recibiovacuna=True)
                        # if 'recibiovacuna' in request.POST:
                        #     vacunado = True if int(request.POST['recibiovacuna']) == 1 else False
                        #     vacunacion.recibiovacuna = vacunado
                        # if 'recibiodosiscompleta' in request.POST:
                        #     recibiodosiscompleta = True if int(request.POST['recibiodosiscompleta']) == 1 else False
                        #     vacunacion.recibiodosiscompleta = recibiodosiscompleta
                        # if 'deseavacunarse' in request.POST:
                        #     deseavacunarse = True if int(request.POST['deseavacunarse']) == 1 else False
                        #     vacunacion.deseavacunarse = deseavacunarse
                        if 'tipovacuna' in request.POST:
                            vacunacion.tipovacuna = form.cleaned_data['tipovacuna'] if form.cleaned_data[
                                'tipovacuna'] else None
                        # if 'fecha_certificado' in request.POST:
                        #     vacunacion.fecha_certificado = form.cleaned_data['fecha_certificado'] if form.cleaned_data['fecha_certificado'] else None
                        if 'deseavacunarse' in request.POST:
                            vacunacion.deseavacunarse = True if int(request.POST['deseavacunarse']) == 1 else False
                        if 'certificado' in request.FILES:
                            newfile = request.FILES['certificado']
                            extension = newfile._name.split('.')
                            tam = len(extension)
                            exte = extension[tam - 1]
                            if newfile.size > 4194304:
                                return JsonResponse(
                                    {"result": True, "mensaje": u"Error, el tamaño del archivo es mayor a 4 Mb."})
                            if exte in ['pdf', 'jpg', 'jpeg', 'png', 'jpeg', 'peg']:
                                newfile._name = generar_nombre("certificado_", newfile._name)
                            else:
                                return JsonResponse(
                                    {"result": True, "mensaje": u"Error, solo archivos .pdf,.jpg, .jpeg"})
                            vacunacion.certificado = newfile
                        vacunacion.save(request)
                        # if vacunado:
                        datosdosis = request.POST.getlist('infoDosis[]')
                        if datosdosis:
                            c = 0
                            while c < len(datosdosis):
                                dosis = VacunaCovidDosis(cabvacuna=vacunacion, numdosis=datosdosis[c],
                                                         fechadosis=datosdosis[c + 1])
                                dosis.save(request)
                                c += 2
                        else:
                            transaction.set_rollback(True)
                            return JsonResponse({"result": True, "mensaje": "Registre al menos una dosis."},
                                                safe=False)

                        log(u'Adiciono evidencia de vacunación: %s' % vacunacion, request, "add")
                        return JsonResponse({"result": False}, safe=False)
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({"result": True, "mensaje": "Complete los datos requeridos."}, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        elif action == 'delvacunacion':
            try:
                persona = request.session['persona']
                vacuna = VacunaCovid.objects.get(pk=int(request.POST['id']))
                vacuna.status = False
                vacuna.save(request)
                log(u'Elimino Evidencia de Vacunación: %s' % vacuna, request, "del")
                res_js = {'error': False}
            except Exception as ex:
                transaction.set_rollback(True)
                err = 'Error on line {}'.format(sys.exc_info()[-1].tb_lineno)
                msg_err = f"Ocurrio un error: {ex.__str__()}. {err}"
                res_js = {'error': True, 'mensaje': msg_err}
            return JsonResponse(res_js)

        elif action == 'adddosisvacuna':
            try:
                with transaction.atomic():
                    vacunacion = VacunaCovid.objects.get(pk=int(request.POST['id']))
                    if 'recibiodosiscompleta' in request.POST:
                        recibiodosiscompleta = True if int(request.POST['recibiodosiscompleta']) == 1 else False
                        vacunacion.recibiodosiscompleta = recibiodosiscompleta
                    vacunacion.save(request)
                    VacunaCovidDosis.objects.filter(cabvacuna=vacunacion).delete()
                    datosdosis = request.POST.getlist('infoDosis[]')
                    if datosdosis:
                        c = 0
                        while c < len(datosdosis):
                            dosis = VacunaCovidDosis(cabvacuna=vacunacion, numdosis=datosdosis[c],
                                                     fechadosis=datosdosis[c + 1])
                            dosis.save(request)
                            c += 2
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({"result": True, "mensaje": "Registre al menos una dosis."}, safe=False)
                    log(u'Adiciono Certificado de vacunación: %s' % vacunacion, request, "add")
                    return JsonResponse({"result": False}, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        elif action == 'addcertificadovacunacovid':
            try:
                with transaction.atomic():
                    newfile = None
                    form = VacunacionCovidEvidenciaForm(request.POST, request.FILES)
                    if form.is_valid():
                        vacunacion = VacunaCovid.objects.get(pk=encrypt_id(request.POST['id']))
                        if 'certificado' in request.FILES:
                            newfile = request.FILES['certificado']
                            extension = newfile._name.split('.')
                            tam = len(extension)
                            exte = extension[tam - 1]
                            if newfile.size > 4194304:
                                return JsonResponse(
                                    {"result": True, "mensaje": u"Error, el tamaño del archivo es mayor a 4 Mb."})
                            if exte in ['pdf', 'jpg', 'jpeg', 'png', 'jpeg', 'peg']:
                                newfile._name = generar_nombre("certificado_", newfile._name)
                            else:
                                return JsonResponse(
                                    {"result": True, "mensaje": u"Error, solo archivos .pdf,.jpg, .jpeg"})
                            vacunacion.certificado = newfile
                        vacunacion.save(request)
                        log(u'Adiciono Certificado de vacunación: %s' % vacunacion, request, "add")
                        return JsonResponse({"result": False}, safe=False)
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({"result": True, "mensaje": "Complete los datos requeridos."}, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)


        # CAPACITACION

        elif action == 'addcapacitacion':
            try:
                persona = request.session['persona']
                f = CapacitacionPersonaForm(request.POST, request.FILES)
                if 'archivo' in request.FILES:
                    d = request.FILES['archivo']
                    if d.size > 4194304:
                        raise NameError("Error, archivo mayor a 4 Mb.")
                    else:
                        newfiles = request.FILES['archivo']
                        newfilesd = newfiles._name
                        ext = newfilesd[newfilesd.rfind("."):]
                        if ext == '.pdf' or ext == '.PDF' or ext == '.png' or ext == '.jpg' or ext == '.jpeg':
                            tipocapacitacion = int(request.POST['tipocapacitacion'])
                            if tipocapacitacion == 2:
                                f.fields['anio'].required = False
                                anio = datetime.now().date().year
                            if not f.is_valid():
                                transaction.set_rollback(True)
                                return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()],
                                                     "mensaje": "Error en el formulario"})
                            newfile = request.FILES['archivo']
                            newfile._name = generar_nombre("capacitacion_", newfile._name)
                            anio = f.cleaned_data['anio']
                            capacitacion = Capacitacion(persona=persona,
                                                        institucion=f.cleaned_data['institucion'],
                                                        nombre=f.cleaned_data['nombre'],
                                                        descripcion=f.cleaned_data['descripcion'],
                                                        tipo=f.cleaned_data['tipo'],
                                                        tipocurso=f.cleaned_data['tipocurso'],
                                                        tipocapacitacion=f.cleaned_data['tipocapacitacion'],
                                                        tipocertificacion=f.cleaned_data['tipocertificacion'],
                                                        tipoparticipacion=f.cleaned_data['tipoparticipacion'],
                                                        anio=anio,
                                                        contextocapacitacion=f.cleaned_data['contexto'],
                                                        detallecontextocapacitacion=f.cleaned_data['detallecontexto'],
                                                        auspiciante=f.cleaned_data['auspiciante'],
                                                        areaconocimiento=f.cleaned_data['areaconocimiento'],
                                                        subareaconocimiento=f.cleaned_data['subareaconocimiento'],
                                                        subareaespecificaconocimiento=f.cleaned_data['subareaespecificaconocimiento'],
                                                        pais=f.cleaned_data['pais'],
                                                        provincia=f.cleaned_data['provincia'],
                                                        canton=f.cleaned_data['canton'],
                                                        parroquia=f.cleaned_data['parroquia'],
                                                        fechainicio=f.cleaned_data['fechainicio'],
                                                        fechafin=f.cleaned_data['fechafin'],
                                                        horas=f.cleaned_data['horas'],
                                                        expositor=f.cleaned_data['expositor'],
                                                        modalidad=f.cleaned_data['modalidad'],
                                                        otramodalidad=f.cleaned_data['otramodalidad'],
                                                        archivo=newfile)
                            capacitacion.save(request)
                            if DistributivoPersona.objects.filter(status=True, persona=persona):
                                lista = ['talento_humano@unemi.edu.ec', 'portizg@unemi.edu.ec']
                                asunto = "Ingresaron nueva capacitación (archivo)"
                                send_html_mail(asunto, "emails/ingreso_archivos_tthh.html",
                                               {'asunto': asunto,
                                                'd': capacitacion.persona.nombre_completo_inverso(),
                                                'fecha': datetime.now().date(), 'escenario': 'capacitación'}, lista,
                                               [],
                                               cuenta=CUENTAS_CORREOS[1][1])

                            log(u'Adiciono capacitacion: %s' % persona, request, "add")
                            return JsonResponse({'result': False, 'mensaje': 'Guardado con exito'})
                        else:
                            raise NameError("Error, solo archivos .pdf, png, jpg, jpeg.")
                else:
                    raise NameError("No existe archivo")
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': f'{ex}'})

        elif action == 'editcapacitacion':
            try:
                persona = request.session['persona']
                capacitacion = Capacitacion.objects.get(pk=encrypt_id(request.POST['id']))
                if not capacitacion.archivo and not 'archivo' in request.FILES:
                    raise NameError('No ha seleccionado archivo')
                if 'archivo' in request.FILES:
                    d = request.FILES['archivo']
                    if d.size > 4194304:
                        raise NameError("Error, archivo mayor a 4 Mb.")
                    else:
                        newfiles = request.FILES['archivo']
                        newfilesd = newfiles._name
                        ext = newfilesd[newfilesd.rfind("."):]
                        if not (ext == '.pdf' or ext == '.png' or ext == '.jpg' or ext == '.jpeg'):
                            raise NameError("Error, solo archivos .pdf, png, jpg, jpeg.")
                f = CapacitacionPersonaForm(request.POST, request.FILES, instancia=capacitacion)
                tipocapacitacion = int(request.POST['tipocapacitacion'])
                if tipocapacitacion == 2:
                    f.fields['anio'].required = False
                    anio = datetime.now().date().year
                if not f.is_valid():
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()],
                                         "mensaje": "Error en el formulario"})
                anio = f.cleaned_data['anio']
                capacitacion.institucion = f.cleaned_data['institucion']
                capacitacion.nombre = f.cleaned_data['nombre']
                capacitacion.descripcion = f.cleaned_data['descripcion']
                capacitacion.tipo = f.cleaned_data['tipo']
                capacitacion.tipocurso = f.cleaned_data['tipocurso']
                capacitacion.tipocapacitacion = f.cleaned_data['tipocapacitacion']
                capacitacion.tipocertificacion = f.cleaned_data['tipocertificacion']
                capacitacion.tipoparticipacion = f.cleaned_data['tipoparticipacion']
                capacitacion.auspiciante = f.cleaned_data['auspiciante']
                capacitacion.areaconocimiento = f.cleaned_data['areaconocimiento']
                capacitacion.subareaconocimiento = f.cleaned_data['subareaconocimiento']
                capacitacion.subareaespecificaconocimiento = f.cleaned_data['subareaespecificaconocimiento']
                capacitacion.pais = f.cleaned_data['pais']
                capacitacion.anio = anio
                capacitacion.contextocapacitacion = f.cleaned_data['contexto']
                capacitacion.detallecontextocapacitacion = f.cleaned_data['detallecontexto']
                capacitacion.provincia = f.cleaned_data['provincia']
                capacitacion.canton = f.cleaned_data['canton']
                capacitacion.parroquia = f.cleaned_data['parroquia']
                capacitacion.fechainicio = f.cleaned_data['fechainicio']
                capacitacion.fechafin = f.cleaned_data['fechafin']
                capacitacion.horas = f.cleaned_data['horas']
                capacitacion.expositor = f.cleaned_data['expositor']
                capacitacion.modalidad = f.cleaned_data['modalidad']
                capacitacion.otramodalidad = f.cleaned_data['otramodalidad']
                if 'archivo' in request.FILES:
                    newfile = request.FILES['archivo']
                    newfile._name = generar_nombre("capacitacion_", newfile._name)
                    capacitacion.archivo = newfile
                # capacitacion.tiempo = f.cleaned_data['tiempo']
                capacitacion.save(request)
                log(u'Modifoco capacitacion: %s' % persona, request, "edit")
                return JsonResponse({'result': False, 'mensaje':'Guardado con éxito'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': f'{ex}'})

        elif action == 'delcapacitacion':
            try:
                capacitacion = Capacitacion.objects.get(pk=encrypt_id(request.POST['id']))
                if capacitacion.verificado:
                    raise NameError('No puede eliminar la capacitacion.')
                log(u'Elimino capacitacion: %s' % capacitacion, request, "del")
                capacitacion.status=False
                capacitacion.save(request)
                res_js = {'error': False}
            except Exception as ex:
                transaction.set_rollback(True)
                err = 'Error on line {}'.format(sys.exc_info()[-1].tb_lineno)
                msg_err = f"Ocurrio un error: {ex.__str__()}. {err}"
                res_js = {'error': True, 'mensaje': msg_err}
            return JsonResponse(res_js)

        elif action == 'deleventoaprobacion':
            try:
                cabecera = CapCabeceraSolicitud.objects.get(pk=encrypt_id(request.POST['id']))
                cabecera.capdetallesolicitud_set.filter(status=True).update(status=False)
                cabecera.status=False
                cabecera.save(request)
                log(u'Elimino Cabecera Solicitud de Evento: %s' % cabecera, request, "del")
                res_js = {'error': False}
            except Exception as ex:
                transaction.set_rollback(True)
                err = 'Error on line {}'.format(sys.exc_info()[-1].tb_lineno)
                msg_err = f"Ocurrio un error: {ex.__str__()}. {err}"
                res_js = {'error': True, 'mensaje': msg_err}
            return JsonResponse(res_js)

        elif action == 'addcabsolicitud':
            try:
                if Departamento.objects.filter(integrantes__isnull=False, responsable=persona,
                                               status=True).distinct().exists():
                    cabecera = CapCabeceraSolicitud(capeventoperiodo_id=int(request.POST['id']),
                                                    solicita=persona,
                                                    fechasolicitud=datetime.now().date(),
                                                    estadosolicitud=variable_valor('PENDIENTE_CAPACITACION'),
                                                    participante=persona)
                    cabecera.save(request)
                    log(u'Ingreso Cabecera Solicitud de Evento : %s' % cabecera, request, "add")
                    detalle = CapDetalleSolicitud(cabecera=cabecera,
                                                  aprueba=persona,
                                                  observacion="SOLICITUD DE INSCRIPCIÓN",
                                                  fechaaprobacion=datetime.now().date(),
                                                  estado=2)
                    detalle.save(request)
                    detalle.mail_notificar_talento_humano(request.session['nombresistema'], True)
                    log(u'Ingreso Detalle Solicitud de Evento : %s' % detalle, request, "add")
                    res_js = {'error': False}
                else:
                    cabecera = CapCabeceraSolicitud(capeventoperiodo_id=int(request.POST['id']),
                                                    solicita=persona,
                                                    fechasolicitud=datetime.now().date(),
                                                    estadosolicitud=variable_valor('SOLICITUD_CAPACITACION'),
                                                    participante=persona)
                    cabecera.save(request)
                    log(u'Ingreso Cabecera Solicitud de Evento : %s' % cabecera, request, "add")
                    detalle = CapDetalleSolicitud(cabecera=cabecera,
                                                  aprueba=persona,
                                                  observacion="SOLICITUD DE INSCRIPCIÓN",
                                                  fechaaprobacion=datetime.now().date(),
                                                  estado=1)
                    detalle.save(request)
                    detalle.mail_notificar_jefe_departamento(request.session['nombresistema'], False)
                    log(u'Ingreso Detalle Solicitud de Evento : %s' % detalle, request, "add")
                    res_js = {'error': False}
            except Exception as ex:
                transaction.set_rollback(True)
                err = 'Error on line {}'.format(sys.exc_info()[-1].tb_lineno)
                msg_err = f"Ocurrio un error: {ex.__str__()}. {err}"
                res_js = {'error': True, 'mensaje': msg_err}
            return JsonResponse(res_js)

        elif action == 'saveEncuesta':
            from inno.models import CapRespuestaEncuestaSatisfaccion, CapPreguntaEncuestaPeriodo, \
                CapOpcionPreguntaEncuestaPeriodo
            try:
                id = encrypt(request.POST.get('id', encrypt('0')))
                try:
                    eCapCabecera = CapCabeceraSolicitudDocente.objects.get(id=id)
                except ObjectDoesNotExist:
                    raise NameError(u"No se encontro solicitud del evento")
                eCapEvento = eCapCabecera.capeventoperiodo
                eCapEncuesta = eCapEvento.encuesta
                eCapPreguntas = CapPreguntaEncuestaPeriodo.objects.filter(encuesta=eCapEncuesta, status=True, isActivo=True)
                # aRespuestas = []
                for eCapPregunta in eCapPreguntas:
                    opcion_id = request.POST.get(f"opcion_pregunta_id_{encrypt(eCapPregunta.id)}", 0)
                    if opcion_id == '' and type(opcion_id) is str:
                        opcion_id = 0
                    else:
                        opcion_id = int(opcion_id)
                        
                    # valoracion = int(request.POST.get(f"valoracion_id_{encrypt(eCapPregunta.id)}", 0))
                    observacion = request.POST.get(f"observacion_id_{encrypt(eCapPregunta.id)}", None)
                    try:
                        eCapOpcionPreguntaEncuestaPeriodo = CapOpcionPreguntaEncuestaPeriodo.objects.get(id=opcion_id)
                    except ObjectDoesNotExist:
                        raise NameError(f"Seleccione una opción de la pregunta {eCapPregunta.descripcion}")
                    #
                    # if valoracion == 0:
                    #     raise NameError(f"Pregunta {eCapPregunta.descripcion}")
                    try:
                        eCapRespuestaEncuesta = CapRespuestaEncuestaSatisfaccion.objects.get(pregunta=eCapPregunta, solicitud=eCapCabecera)
                    except ObjectDoesNotExist:
                        eCapRespuestaEncuesta = CapRespuestaEncuestaSatisfaccion(pregunta=eCapPregunta,
                                                                                 solicitud=eCapCabecera)
                    eCapRespuestaEncuesta.opcion = eCapOpcionPreguntaEncuestaPeriodo
                    eCapRespuestaEncuesta.valoracion = eCapOpcionPreguntaEncuestaPeriodo.valoracion
                    eCapRespuestaEncuesta.observacion = observacion
                    eCapRespuestaEncuesta.save(request)
                    log(u'Respuesta de encuesta de satisfacción : %s' % eCapRespuestaEncuesta, request, "add")
                res_js = {'isSuccess': True,
                          'message': u"Se guardo correctamente la encuesta.",
                          'data': {'url': eCapCabecera.rutapdf.url if eCapCabecera.rutapdf else None,
                                   'id': eCapCabecera.id,
                                   'name': eCapCabecera.rutapdf.name if eCapCabecera.rutapdf else None}}
            except Exception as ex:
                transaction.set_rollback(True)
                err = 'Error on line {}'.format(sys.exc_info()[-1].tb_lineno)
                if DEBUG:
                    print(err)
                msg_err = f"{ex.__str__()}"
                res_js = {'isSuccess': False,
                          'message': msg_err,
                          'data': {'url': None, 'id': None, 'name': None}}
            return JsonResponse(res_js)

        # CERTIFICACIONES

        elif action == 'addcertificadoidioma':
            try:
                persona = request.session['persona']
                f = CertificadoIdiomaForm(request.POST, request.FILES)
                if 'archivo' in request.FILES:
                    d = request.FILES['archivo']
                    newfilesd = d._name
                    ext = newfilesd[newfilesd.rfind("."):]
                    if ext == '.pdf':
                        a = 1
                    else:
                        return JsonResponse({"result": "bad", "mensaje": u"Error, solo archivos .pdf."})
                    if d.size > 2194304:
                        return JsonResponse({"result": "bad", "mensaje": u"Error, archivo mayor a 2 Mb."})
                if not f.is_valid():
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()],
                                         "mensaje": "Error en el formulario"})
                certificado = CertificadoIdioma(persona=persona,
                                                idioma=f.cleaned_data['idioma'],
                                                institucioncerti=f.cleaned_data['institucioncerti'],
                                                validainst=f.cleaned_data['validainst'],
                                                otrainstitucion=f.cleaned_data['otrainstitucion'],
                                                nivelsuficencia=f.cleaned_data['nivelsuficencia'],
                                                fechacerti=f.cleaned_data['fechacerti'])
                certificado.save(request)
                newfile = None
                if 'archivo' in request.FILES:
                    newfile = request.FILES['archivo']
                    newfile._name = generar_nombre("certificado_otro_", newfile._name)
                    certificado.archivo = newfile
                    certificado.save(request)
                log(u'Adiciono certificado internacional: %s' % persona, request, "add")
                return JsonResponse({'result': False, 'mensaje': 'Guardado con exito'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': f'Error al guardar los datos: {ex}'})

        elif action == 'editcertificadoidioma':
            try:
                persona = request.session['persona']
                certificado = CertificadoIdioma.objects.get(pk=encrypt_id(request.POST['id']))
                if 'archivo' in request.FILES:
                    d = request.FILES['archivo']
                    if d.size > 2194304:
                        return JsonResponse({"result": "bad", "mensaje": u"Error, archivo mayor a 2 Mb."})
                    else:
                        newfiles = request.FILES['archivo']
                        newfilesd = newfiles._name
                        ext = newfilesd[newfilesd.rfind("."):]
                        if not (ext == '.pdf'):
                            return JsonResponse({"result": "bad", "mensaje": u"Error, solo archivos .pdf"})
                f = CertificadoIdiomaForm(request.POST, request.FILES)
                if not f.is_valid():
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()],
                                         "mensaje": "Error en el formulario"})
                certificado.idioma = f.cleaned_data['idioma']
                certificado.institucioncerti = f.cleaned_data['institucioncerti']
                certificado.otrainstitucion = f.cleaned_data['otrainstitucion']
                certificado.nivelsuficencia = f.cleaned_data['nivelsuficencia']
                certificado.fechacerti = f.cleaned_data['fechacerti']
                certificado.validainst = f.cleaned_data['validainst']
                if 'archivo' in request.FILES:
                    newfile = request.FILES['archivo']
                    newfile._name = generar_nombre("certificado_otro_", newfile._name)
                    certificado.archivo = newfile
                certificado.save(request)
                log(u'Modifico certificación de idioma personal: %s' % persona, request, "edit")
                return JsonResponse({'result': False, 'mensaje': 'Guardado con exito'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': f'Error al guardar los datos: {ex}'})

        elif action == 'delcertificadoidioma':
            try:
                certificadopersona = CertificadoIdioma.objects.get(pk=encrypt_id(request.POST['id']))
                certificadopersona.status = False
                certificadopersona.save(request)
                log(u'Eliminó un certificado internacional personal a la hoja de vida: %s - la persona: %s' % (certificadopersona, persona), request, "add")
                res_js = {'error': False}
            except Exception as ex:
                transaction.set_rollback(True)
                err = 'Error on line {}'.format(sys.exc_info()[-1].tb_lineno)
                msg_err = f"Ocurrio un error: {ex.__str__()}. {err}"
                res_js = {'error': True, 'mensaje': msg_err}
            return JsonResponse(res_js)

        elif action == 'addcertificado':
            try:
                persona = request.session['persona']
                f = CertificadoPersonaForm(request.POST, request.FILES)
                if 'archivo' in request.FILES:
                    d = request.FILES['archivo']
                    newfilesd = d._name
                    ext = newfilesd[newfilesd.rfind("."):]
                    if ext == '.pdf':
                        a = 1
                    else:
                        return JsonResponse({"result": "bad", "mensaje": u"Error, solo archivos .pdf."})
                    if d.size > 2194304:
                        return JsonResponse({"result": "bad", "mensaje": u"Error, archivo mayor a 2 Mb."})
                if not f.is_valid():
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()],
                                         "mensaje": "Error en el formulario"})
                certificado = CertificacionPersona(persona=persona,
                                                   nombres=f.cleaned_data['nombres'],
                                                   autoridad_emisora=f.cleaned_data['autoridad_emisora'],
                                                   numerolicencia=f.cleaned_data['numerolicencia'],
                                                   enlace=f.cleaned_data['enlace'],
                                                   vigente=f.cleaned_data['vigente'],
                                                   fechadesde=f.cleaned_data['fechadesde'],
                                                   fechahasta=f.cleaned_data['fechahasta'])
                certificado.save(request)
                newfile = None
                if 'archivo' in request.FILES:
                    newfile = request.FILES['archivo']
                    newfile._name = generar_nombre("certificado_", newfile._name)
                    certificado.archivo = newfile
                    certificado.save(request)
                log(u'Adiciono certificado: %s' % persona, request, "add")
                return JsonResponse({'result': False, 'mensaje': 'Guardado con exito'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': f'Error al guardar los datos {ex}'})

        elif action == 'editcertificado':
            try:
                persona = request.session['persona']
                certificado = CertificacionPersona.objects.get(pk=encrypt_id(request.POST['id']))
                if 'archivo' in request.FILES:
                    d = request.FILES['archivo']
                    if d.size > 2194304:
                        return JsonResponse({"result": "bad", "mensaje": u"Error, archivo mayor a 2 Mb."})
                    else:
                        newfiles = request.FILES['archivo']
                        newfilesd = newfiles._name
                        ext = newfilesd[newfilesd.rfind("."):]
                        if not (ext == '.pdf'):
                            return JsonResponse({"result": "bad", "mensaje": u"Error, solo archivos .pdf"})
                f = CertificadoPersonaForm(request.POST, request.FILES)
                if not f.is_valid():
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()],
                                         "mensaje": "Error en el formulario"})
                certificado.nombres = f.cleaned_data['nombres']
                certificado.autoridad_emisora = f.cleaned_data['autoridad_emisora']
                certificado.numerolicencia = f.cleaned_data['numerolicencia']
                certificado.fechadesde = f.cleaned_data['fechadesde']
                certificado.fechahasta = f.cleaned_data['fechahasta']
                certificado.enlace = f.cleaned_data['enlace']
                certificado.vigente = f.cleaned_data['vigente']
                if 'archivo' in request.FILES:
                    newfile = request.FILES['archivo']
                    newfile._name = generar_nombre("capacitacion_", newfile._name)
                    certificado.archivo = newfile
                # capacitacion.tiempo = f.cleaned_data['tiempo']
                certificado.save(request)
                log(u'Modifico certificación personal: %s' % persona, request, "edit")
                return JsonResponse({'result': False, 'mensaje': 'Guardado con exito'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': f'Error al guardar los datos {ex}'})

        elif action == 'delcertificado':
            try:
                certificadopersona = CertificacionPersona.objects.get(pk=encrypt_id(request.POST['id']))
                certificadopersona.status = False
                certificadopersona.save(request)
                log(u'Eliminó un certificado personal a la hoja de vida: %s - la persona: %s' % (certificadopersona, persona), request, "add")
                res_js = {'error': False}
            except Exception as ex:
                transaction.set_rollback(True)
                err = 'Error on line {}'.format(sys.exc_info()[-1].tb_lineno)
                msg_err = f"Ocurrio un error: {ex.__str__()}. {err}"
                res_js = {'error': True, 'mensaje': msg_err}
            return JsonResponse(res_js)

        # DISCAPACIDAD

        elif action == 'editdiscapacidad':
            try:
                if 'archivo' in request.FILES:
                    arch = request.FILES['archivo']
                    extension = arch._name.split('.')
                    tam = len(extension)
                    exte = extension[tam - 1]
                    if arch.size > 4194304:
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"Error, el tamaño del archivo es mayor a 4 Mb."})
                    if not exte.lower() in ['pdf']:
                        return JsonResponse({"result": "bad", "mensaje": u"Solo se permiten archivos .pdf"})
                if 'archivovaloracion' in request.FILES:
                    arch = request.FILES['archivovaloracion']
                    extension = arch._name.split('.')
                    tam = len(extension)
                    exte = extension[tam - 1]
                    if arch.size > 4194304:
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"Error, el tamaño del archivo es mayor a 4 Mb."})
                    if not exte.lower() in ['pdf']:
                        return JsonResponse({"result": "bad", "mensaje": u"Solo se permiten archivos .pdf"})

                persona = request.session['persona']
                f = DiscapacidadForm(request.POST, request.FILES)
                if not f.is_valid():
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()],
                                         "mensaje": "Error en el formulario"})
                newfile = None
                if not f.cleaned_data['tienediscapacidad'] and f.cleaned_data['tienediscapacidadmultiple']:
                    raise NameError('No puede marcar discapacidad multiple sin marcar que tiene discapacidad')
                if not f.cleaned_data['tipodiscapacidad'] and f.cleaned_data['tipodiscapacidadmultiple']:
                    raise NameError(
                        'No puede elegir discapacidades multiples sin elegir una discapacidad principal')
                if f.cleaned_data['tienediscapacidadmultiple'] and not f.cleaned_data['tipodiscapacidadmultiple']:
                    raise NameError('Debe elegir una o más discapacidades multiples')
                perfil = persona.mi_perfil()
                perfil.tienediscapacidad = f.cleaned_data['tienediscapacidad']
                perfil.tipodiscapacidad = f.cleaned_data['tipodiscapacidad']
                perfil.porcientodiscapacidad = f.cleaned_data['porcientodiscapacidad'] if f.cleaned_data[
                    'porcientodiscapacidad'] else 0
                perfil.carnetdiscapacidad = f.cleaned_data['carnetdiscapacidad']
                perfil.institucionvalida = f.cleaned_data['institucionvalida']
                perfil.tienediscapacidadmultiple = f.cleaned_data['tienediscapacidadmultiple']
                perfil.grado = f.cleaned_data['grado'] if f.cleaned_data['grado'] else 0

                if not f.cleaned_data['tienediscapacidad']:
                    perfil.archivo = None
                    perfil.estadoarchivodiscapacidad = None
                    perfil.archivovaloracion = None
                if 'archivo' in request.FILES:
                    newfile = request.FILES['archivo']
                    newfile._name = generar_nombre("archivosdiscapacidad_", newfile._name)
                    perfil.archivo = newfile
                    perfil.estadoarchivodiscapacidad = 1
                if 'archivovaloracion' in request.FILES:
                    newfile = request.FILES['archivovaloracion']
                    newfile._name = generar_nombre("archivovaloracionmedica_", newfile._name)
                    perfil.archivovaloracion = newfile
                perfil.save(request)
                perfil.tipodiscapacidadmultiple.clear()
                perfil.subtipodiscapacidad.clear()
                if f.cleaned_data['tienediscapacidadmultiple']:
                    tipos = request.POST.getlist('tipodiscapacidadmultiple')
                    for tipo in tipos:
                        perfil.tipodiscapacidadmultiple.add(tipo)
                if f.cleaned_data['tipodiscapacidad'] and f.cleaned_data['tienediscapacidad']:
                    subtipos = request.POST.getlist('subtipodiscapacidad')
                    for subtipo in subtipos:
                        perfil.subtipodiscapacidad.add(subtipo)
                log(u'Modifico tipo de discapacidad: %s' % persona, request, "edit")
                return JsonResponse({'result': False, 'mensaje': 'Guardado con exito'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': f'Error al guardar los datos {ex}'})

        # BITACORA

        elif action == 'delbitacora':
            try:
                bitacora = BitacoraActividadDiaria.objects.get(pk=int(encrypt(request.POST['id'])))
                log(u'Elimino bitacora de actividad diaria %s' % (bitacora), request, "del")
                bitacora.status = False
                bitacora.save(request)
                res_js = {'error': False}
            except Exception as ex:
                transaction.set_rollback(True)
                err = 'Error on line {}'.format(sys.exc_info()[-1].tb_lineno)
                msg_err = f"Ocurrio un error: {ex.__str__()}. {err}"
                res_js = {'error': True, 'mensaje': msg_err}
            return JsonResponse(res_js)

        elif action == 'addbitacora':
            try:
                persona = request.session['persona']
                form = BitacoraForm(request.POST)
                form.ocultarcampos_tipoactividad()
                form.ocultarcampos_horafin()

                hi = request.POST['hora'] if 'hora' in request.POST else ''
                hf = request.POST['horafin'] if 'horafin' in request.POST else ''

                if not hi:
                    return JsonResponse({"result": "bad", "mensaje": "Seleccione la hora de inicio de la actividad."})

                if hi and hf:
                    h1 = timedelta(hours=int(hi.split(':')[0]), minutes=int(hi.split(':')[1]))
                    h2 = timedelta(hours=int(hf.split(':')[0]), minutes=int(hf.split(':')[1]))
                    total = f"{h2 - h1}"
                    if int(total.split(':')[0]) == 0 and int(total.split(':')[1]) < 29:
                        raise NameError(f"La duración de la actividad debe ser mayor a 29 minutos.")
                # if persona.contratodip_set.filter(status=True).exists() and not persona.es_coordinadorcarrera(request.session['periodo']) and not persona.contratodip_set.filter(cargo__nombre__icontains='TUTOR').exists():
                #     fecha = datetime.today()
                #     fecbitaco = convertir_fecha_hora(u"%s %s" % (request.POST['fecha'], request.POST['hora']))
                #     if datetime.fromisoformat('2022-07-01').date() != fecbitaco.date():
                #         if (fecha - fecbitaco).days > 2:
                #             return JsonResponse({'result': 'bad', 'mensaje': "Excedió fecha limite para subir bitácora "})
                newfile = None
                if 'archivo' in request.FILES:
                    newfile = request.FILES['archivo']
                    if newfile:
                        newfilesd = newfile._name
                        ext = newfilesd[newfilesd.rfind("."):]
                        if newfile.size > 12582912:
                            return JsonResponse({"result": "bad", "mensaje": u"Error, archivo mayor a 12 Mb."})
                        if newfile:
                            newfile._name = generar_nombre("bitacora", newfile._name)

                fechahorafin = convertir_fecha_hora_invertida(u"%s %s" % (request.POST['fecha'], hf)) if hf else None
                if form.is_valid():
                    fecha_form = convertir_fecha_hora_invertida(u"%s %s" % (request.POST['fecha'], request.POST['hora']))
                    bitacora = BitacoraActividadDiaria(persona=persona,
                                                       fecha=fecha_form,
                                                       fechafin=fechahorafin,
                                                       departamento=persona.mi_departamento(),
                                                       descripcion=u'%s' % form.cleaned_data['descripcion'],
                                                       link=form.cleaned_data['link'],
                                                       tiposistema=form.cleaned_data['tiposistema'],
                                                       archivo=newfile,
                                                       departamento_requiriente=form.cleaned_data[
                                                           'departamento_requiriente']
                                                       )
                    if persona.contratodip_set.filter(status=True).exists():
                        contratodip = persona.contratodip_set.filter(status=True).last()
                        hoy = datetime.now().date()
                        if not Persona.objects.filter(id=persona.id,usuario__groups__id=49).exists():
                            if contratodip.fechaaplazo:
                                if contratodip.fechaaplazo < hoy:
                                    fecha_comienza = variable_valor("VALIDAR_FECHA_LIMITE_POSGRADO")
                                    if fecha_form.date() >= fecha_comienza:
                                        valida_fecha = validar_dia_maximo_bitacora(fecha_form,5)
                                        if not valida_fecha[0]:
                                            raise NameError(f"Fecha límite para registrar actividad ha sido superada.\nCon la fecha seleccionada solo tiene 5 días adicionales para registrar la actividad.\nTenía hasta el {valida_fecha[1].date()}")
                            else:
                                fecha_comienza = variable_valor("VALIDAR_FECHA_LIMITE_POSGRADO")
                                if fecha_form.date() >= fecha_comienza:
                                    valida_fecha = validar_dia_maximo_bitacora(fecha_form,5)
                                    if not valida_fecha[0]:
                                        raise NameError( f"Fecha límite para registrar actividad ha sido superada.\nCon la fecha seleccionada solo tiene 5 días adicionales para registrar la actividad.\nTenía hasta el {valida_fecha[1].date()}")
                        bitacora.actividades = form.cleaned_data['actividades']
                        bitacora.tipoactividad = request.POST[
                            'tipoactividad'] if 'tipoactividad' in request.POST else None
                    else:
                        bitacora.titulo = form.cleaned_data['titulo']
                    bitacora.save(request)
                    log(u'Adicionó bitacora de actividades diarias %s' % (bitacora), request, "add")
                    return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos. %s"%(ex.__str__())})

        elif action == 'editbitacora':
            try:
                persona = request.session['persona']
                form = BitacoraForm(request.POST)
                form.ocultarcampos_tipoactividad()
                form.ocultarcampos_horafin()

                hi = request.POST['hora'] if 'hora' in request.POST else ''
                hf = request.POST['horafin'] if 'horafin' in request.POST else ''

                if not hi:
                    return JsonResponse({"result": "bad", "mensaje": "Seleccione la hora de inicio de la actividad."})

                if hi and hf:
                    h1 = timedelta(hours=int(hi.split(':')[0]), minutes=int(hi.split(':')[1]))
                    h2 = timedelta(hours=int(hf.split(':')[0]), minutes=int(hf.split(':')[1]))
                    total = f"{h2 - h1}"
                    if int(total.split(':')[0]) == 0 and int(total.split(':')[1]) < 59:
                        return JsonResponse(
                            {"result": "bad", "mensaje": "La duración de la actividad debe ser mayor a 59 minutos."})

                newfile = None
                if 'archivo' in request.FILES:
                    newfile = request.FILES['archivo']
                    if newfile:
                        newfilesd = newfile._name
                        ext = newfilesd[newfilesd.rfind("."):]
                        if newfile.size > 12582912:
                            return JsonResponse({"result": "bad", "mensaje": u"Error, archivo mayor a 12 Mb."})
                        if newfile:
                            newfile._name = generar_nombre("bitacora", newfile._name)

                fechahorafin = convertir_fecha_hora_invertida(u"%s %s" % (request.POST['fecha'], hf)) if hf else None
                if form.is_valid():
                    fecha_form = convertir_fecha_hora_invertida(u"%s %s" % (request.POST['fecha'], request.POST['hora']))
                    bitacora = BitacoraActividadDiaria.objects.get(pk=request.POST['id'])
                    bitacora.titulo = form.cleaned_data['titulo']
                    bitacora.fecha = fecha_form
                    bitacora.fechafin = fechahorafin
                    bitacora.departamento = persona.mi_departamento()
                    bitacora.descripcion = form.cleaned_data['descripcion']
                    bitacora.link = form.cleaned_data['link']
                    bitacora.tiposistema = form.cleaned_data['tiposistema']
                    bitacora.departamento_requiriente = form.cleaned_data['departamento_requiriente']
                    if persona.contratodip_set.filter(status=True).exists():
                        contratodip = persona.contratodip_set.filter(status=True).last()
                        hoy = datetime.now().date()
                        # if contratodip.fechaaplazo:
                        #     if contratodip.fechaaplazo < hoy:
                        #         fecha_comienza = variable_valor("VALIDAR_FECHA_LIMITE_POSGRADO")
                        #         if fecha_form.date() >= fecha_comienza:
                        #             valida_fecha = validar_dia_maximo_bitacora(fecha_form,4)
                        #             if not valida_fecha[0]:
                        #                 raise NameError(f"Fecha límite para registrar actividad ha sido superada.\nCon la fecha seleccionada solo tiene 2 días adicionales para registrar la actividad.\nTenía hasta el {valida_fecha[1].date()}")
                        # else:
                        #     fecha_comienza = variable_valor("VALIDAR_FECHA_LIMITE_POSGRADO")
                        #     if fecha_form.date() >= fecha_comienza:
                        #         valida_fecha = validar_dia_maximo_bitacora(fecha_form,4)
                        #         if not valida_fecha[0]:
                        #             raise NameError(f"Fecha límite para registrar actividad ha sido superada.\nCon la fecha seleccionada solo tiene 2 días adicionales para registrar la actividad.\nTenía hasta el {valida_fecha[1].date()}")
                        bitacora.actividades = form.cleaned_data['actividades']
                        bitacora.tipoactividad = request.POST[
                            'tipoactividad'] if 'tipoactividad' in request.POST else None
                    else:
                        bitacora.titulo = form.cleaned_data['titulo']
                    if newfile:
                        bitacora.archivo = newfile
                    bitacora.save(request)
                    log(u'Editó bitacora de actividad diaria %s' % (bitacora), request, "edit")
                    return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos. %s"%(ex.__str__())})

        elif action == 'deletebitacora':
            try:
                with transaction.atomic():
                    instancia = InformeActividadJornada.objects.get(pk=int(request.POST['id']))
                    instancia.status = False
                    instancia.save(request)
                    log(u'Elimino informe de actividades y jornada laboral: %s %s' % (
                        instancia, instancia.contrato.persona), request, "delete")
                    res_json = {"error": False}
            except Exception as ex:
                res_json = {'error': True, "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)

        # FORMACIÓN ACADÉMICA

        elif action == 'addtitulacion':
            try:
                persona = request.session['persona']
                f = TitulacionPersonaForm(request.POST, request.FILES)
                if 'archivo' in request.FILES:
                    arch = request.FILES['archivo']
                    exte = ext_archive(arch._name)
                    if arch.size > 4194304:
                        return JsonResponse({"result": "bad", "mensaje": u"Error, archivo mayor a 4 Mb."})
                    if not exte in  ['.pdf', '.png','.jpg', '.jpeg','.jpg']:
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"Archivo Título solo en formato .pdf, jpg, jpeg, png"})
                if 'registroarchivo' in request.FILES:
                    registroarchivo = request.FILES['registroarchivo']
                    exte1 = ext_archive(registroarchivo._name)
                    if registroarchivo.size > 4194304:
                        return JsonResponse({"result": "bad", "mensaje": u"Error, archivo mayor a 4 Mb."})
                    if not exte1 in  ['.pdf', '.png','.jpg', '.jpeg','.jpg']:
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"Archivo SENESCYT solo en formato .pdf, jpg, jpeg, png"})
                if not f.is_valid():
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()],
                                         "mensaje": "Error en el formulario"})
                titulacion = Titulacion(persona=persona,
                                        titulo=f.cleaned_data['titulo'],
                                        areatitulo=f.cleaned_data['areatitulo'],
                                        fechainicio=f.cleaned_data['fechainicio'],
                                        fechaobtencion=f.cleaned_data['fechaobtencion'],
                                        fechaegresado=f.cleaned_data['fechaegresado'],
                                        registro=f.cleaned_data['registro'],
                                        # areaconocimiento=f.cleaned_data['areaconocimiento'],
                                        # subareaconocimiento=f.cleaned_data['subareaconocimiento'],
                                        # subareaespecificaconocimiento=f.cleaned_data['subareaespecificaconocimiento'],
                                        pais=f.cleaned_data['pais'],
                                        provincia=f.cleaned_data['provincia'],
                                        canton=f.cleaned_data['canton'],
                                        parroquia=f.cleaned_data['parroquia'],
                                        educacionsuperior=f.cleaned_data['educacionsuperior'],
                                        institucion=f.cleaned_data['institucion'],
                                        colegio=f.cleaned_data['colegio'],
                                        anios=f.cleaned_data['anios'],
                                        semestres=f.cleaned_data['semestres'],
                                        cursando=f.cleaned_data['cursando'],
                                        aplicobeca=f.cleaned_data['aplicobeca'],
                                        tipobeca=f.cleaned_data['tipobeca'] if f.cleaned_data[
                                            'aplicobeca'] else None,
                                        financiamientobeca=f.cleaned_data['financiamientobeca'] if f.cleaned_data[
                                            'aplicobeca'] else None,
                                        valorbeca=f.cleaned_data['valorbeca'] if f.cleaned_data[
                                            'aplicobeca'] else 0)
                titulacion.save(request)
                if not titulacion.cursando:
                    if 'archivo' in request.FILES:
                        newfile = request.FILES['archivo']
                        if newfile:
                            newfile._name = generar_nombre("titulacion_", newfile._name)
                            titulacion.archivo = newfile
                            titulacion.save(request)
                        if DistributivoPersona.objects.filter(status=True, persona=persona):
                            lista = ['talento_humano@unemi.edu.ec', 'portizg@unemi.edu.ec']
                            asunto = "Ingresaron título académico (archivo)"
                            send_html_mail(asunto, "emails/ingreso_archivos_tthh.html",
                                           {'asunto': asunto, 'd': titulacion.persona.nombre_completo_inverso(),
                                            'fecha': datetime.now().date(), 'escenario': 'titulación - SENECYT'},
                                           lista, [], cuenta=CUENTAS_CORREOS[1][1])

                    if 'registroarchivo' in request.FILES:
                        newfile2 = request.FILES['registroarchivo']
                        if newfile2:
                            newfile2._name = generar_nombre("archivosenecyt_", newfile2._name)
                            titulacion.registroarchivo = newfile2
                            titulacion.save(request)
                        if DistributivoPersona.objects.filter(status=True, persona=persona):
                            lista = ['talento_humano@unemi.edu.ec', 'portizg@unemi.edu.ec']
                            asunto = "Ingresaron título académico (archivo SENESCYT)"
                            send_html_mail(asunto, "emails/ingreso_archivos_tthh.html",
                                           {'asunto': asunto, 'd': titulacion.persona.nombre_completo_inverso(),
                                            'fecha': datetime.now().date(),
                                            'escenario': 'titulación - título academico'}, lista, [],
                                           cuenta=CUENTAS_CORREOS[1][1])
                # campotitulo = None
                # if CamposTitulosPostulacion.objects.filter(status=True, titulo=f.cleaned_data['titulo']).exists():
                #     campotitulo = CamposTitulosPostulacion.objects.filter(status=True,
                #                                                           titulo=f.cleaned_data['titulo']).first()
                # else:
                #     campotitulo = CamposTitulosPostulacion(titulo=f.cleaned_data['titulo'])
                #     campotitulo.save(request)
                # for ca in f.cleaned_data['campoamplio']:
                #     if not campotitulo.campoamplio.filter(id=ca.id):
                #         campotitulo.campoamplio.add(ca)
                # for ce in f.cleaned_data['campoespecifico']:
                #     if not campotitulo.campoespecifico.filter(id=ce.id):
                #         campotitulo.campoespecifico.add(ce)
                # for cd in f.cleaned_data['campodetallado']:
                #     if not campotitulo.campodetallado.filter(id=cd.id):
                #         campotitulo.campodetallado.add(cd)
                # campotitulo.save()
                log(u'Adiciono titulacion: %s' % persona, request, "add")
                return JsonResponse({'result': False, 'mensaje': 'Guardado con exito'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': f'Error al guardar los datos {ex}'})

        elif action == 'edittitulacion':
            try:
                persona = request.session['persona']
                titulacion = Titulacion.objects.get(pk=encrypt_id(request.POST['id']))
                f = TitulacionPersonaForm(request.POST, request.FILES)
                if titulacion.verificadosenescyt:
                    f.fields['titulo'].required = False

                if 'archivo' in request.FILES:
                    arch = request.FILES['archivo']
                    extencion = arch._name.split('.')
                    exte = extencion[1]
                    if arch.size > 4194304:
                        return JsonResponse({"result": "bad", "mensaje": u"Error, archivo mayor a 4 Mb."})
                    if not exte == 'pdf' and not exte == 'png' and not exte == 'jpg' and not exte == 'jpeg' and not exte == 'jpg':
                        return JsonResponse({"result": "bad", "mensaje": u"Solo archivos .pdf, jpg, jpeg, png"})

                if 'registroarchivo' in request.FILES:
                    registroarchivo = request.FILES['registroarchivo']
                    extencion1 = registroarchivo._name.split('.')
                    exte1 = extencion1[1]
                    if registroarchivo.size > 4194304:
                        return JsonResponse({"result": "bad", "mensaje": u"Error, archivo mayor a 4 Mb."})
                    if not exte1 == 'pdf' and not exte1 == 'png' and not exte1 == 'jpg' and not exte1 == 'jpeg' and not exte1 == 'jpg':
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"Archivo SENESCYT solo en formato .pdf, jpg, jpeg, png"})
                if not f.is_valid():
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()],
                                         "mensaje": "Error en el formulario"})

                titulacion.areatitulo = f.cleaned_data['areatitulo']
                titulacion.fechainicio = f.cleaned_data['fechainicio']
                titulacion.fechaobtencion = f.cleaned_data['fechaobtencion']
                titulacion.fechaegresado = f.cleaned_data['fechaegresado']
                # titulacion.areaconocimiento = f.cleaned_data['areaconocimiento']
                # titulacion.subareaconocimiento = f.cleaned_data['subareaconocimiento']
                # titulacion.subareaespecificaconocimiento = f.cleaned_data['subareaespecificaconocimiento']
                titulacion.pais = f.cleaned_data['pais']
                titulacion.provincia = f.cleaned_data['provincia']
                titulacion.canton = f.cleaned_data['canton']
                titulacion.parroquia = f.cleaned_data['parroquia']
                titulacion.colegio = f.cleaned_data['colegio']
                titulacion.anios = f.cleaned_data['anios']
                titulacion.semestres = f.cleaned_data['semestres']
                titulacion.aplicobeca = f.cleaned_data['aplicobeca']
                titulacion.tipobeca = f.cleaned_data['tipobeca'] if f.cleaned_data['aplicobeca'] else None
                titulacion.financiamientobeca = f.cleaned_data['financiamientobeca'] if f.cleaned_data[
                    'aplicobeca'] else None
                titulacion.valorbeca = f.cleaned_data['valorbeca'] if f.cleaned_data['aplicobeca'] else 0
                if not titulacion.verificadosenescyt:
                    titulacion.titulo = f.cleaned_data['titulo']
                    titulacion.educacionsuperior = f.cleaned_data['educacionsuperior']
                    titulacion.institucion = f.cleaned_data['institucion']
                    titulacion.registro = f.cleaned_data['registro']
                    titulacion.cursando = f.cleaned_data['cursando']
                if 'archivo' in request.FILES:
                    newfile = request.FILES['archivo']
                    newfile._name = generar_nombre("titulacion_", newfile._name)
                    titulacion.archivo = newfile
                if 'registroarchivo' in request.FILES:
                    newfile2 = request.FILES['registroarchivo']
                    if newfile2:
                        newfile2._name = generar_nombre("archivosenecyt_", newfile2._name)
                        titulacion.registroarchivo = newfile2
                titulacion.save(request)
                # campotitulo = None
                # if CamposTitulosPostulacion.objects.filter(status=True, titulo=f.cleaned_data['titulo']).exists():
                #     campotitulo = CamposTitulosPostulacion.objects.filter(status=True,
                #                                                           titulo=f.cleaned_data['titulo']).first()
                # else:
                #     campotitulo = CamposTitulosPostulacion(titulo=f.cleaned_data['titulo'])
                #     campotitulo.save(request)
                # for ca in f.cleaned_data['campoamplio']:
                #     if not campotitulo.campoamplio.filter(id=ca.id):
                #         campotitulo.campoamplio.add(ca)
                # for ce in f.cleaned_data['campoespecifico']:
                #     if not campotitulo.campoespecifico.filter(id=ce.id):
                #         campotitulo.campoespecifico.add(ce)
                # for cd in f.cleaned_data['campodetallado']:
                #     if not campotitulo.campodetallado.filter(id=cd.id):
                #         campotitulo.campodetallado.add(cd)
                # campotitulo.save()
                request.session['instruccion'] = 1
                if Graduado.objects.filter(status=True, inscripcion__persona__id=persona.id).exists():
                    datos = Persona.objects.get(status=True, id=persona.id)
                    if 'personales' in request.session and 'nacimiento' in request.session and 'domicilio' in request.session and 'etnia' in request.session and 'instruccion' in request.session and datos.datosactualizados == 0:
                        if request.session['personales'] == 1 and request.session['nacimiento'] == 1 and \
                                request.session['domicilio'] == 1 and request.session['etnia'] == 1 and \
                                request.session['instruccion'] == 1 and datos.datosactualizados == 0:
                            datos.datosactualizados = 1
                            datos.save(request)
                log(u'Modifico titulacion: %s' % persona, request, "edit")
                return JsonResponse({'result': False, 'mensaje': 'Guardado con exito'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': f'Error al guardar los datos {ex}'})

        elif action == 'deltitulacion':
            try:
                persona = request.session['persona']
                titulacion = Titulacion.objects.get(pk=encrypt_id(request.POST['id']))
                if titulacion.verificado:
                    return JsonResponse({'result': 'bad', 'mensaje': u'No puede eliminar el titulo.'})
                log(u'Elimino titulacion: %s' % titulacion, request, "del")
                titulacion.status = False
                titulacion.save()
                res_js = {'error': False}
            except Exception as ex:
                transaction.set_rollback(True)
                err = 'Error on line {}'.format(sys.exc_info()[-1].tb_lineno)
                msg_err = f"Ocurrio un error: {ex.__str__()}. {err}"
                res_js = {'error': True, 'mensaje': msg_err}
            return JsonResponse(res_js)

        elif action == 'tituloprincipal':
            try:
                titulacion = Titulacion.objects.get(pk=encrypt_id(request.POST['id']))
                val=eval(request.POST['val'].capitalize())
                if val:
                    # if not titulacion.verificado:
                    #     return JsonResponse({'result': 'bad', 'mensaje': u'No se ha verificado su titulo'})
                    titulacion.persona.titulacion_set.update(principal=False)
                    titulacion.principal = val
                else:
                    titulacion.principal = val
                titulacion.save(request)
                log(u"Selecciono titulo principal: %s" % titulacion, request, "del")
                return JsonResponse({'result': True, 'mensaje': 'Guardado con exito'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': False, 'mensaje': f'{ex}'})

        elif action == 'adddetalletitulobachiller':
            try:
                f = DetalleTitulacionBachillerForm(request.POST, request.FILES)

                if not f.is_valid():
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()],
                                         "mensaje": "Error en el formulario"})
                titulacion = Titulacion.objects.get(pk=encrypt_id(request.POST['id']))
                newfile = None
                newfile2 = None
                if 'archivo' in request.FILES:
                    newfile = request.FILES['archivo']
                    if newfile:
                        newfile._name = generar_nombre("actagradobachiller_", newfile._name)
                # else:
                #     h = Truncator(request.POST['archivo1']).chars(-30, truncate = '/')
                #     newfile = f.cleaned_data['archivo1']

                if 'reconocimientoacademico' in request.FILES:
                    newfile2 = request.FILES['reconocimientoacademico']

                    if newfile2:
                        newfile2._name = generar_nombre("reconocimientoacademico_", newfile2._name)
                # else:
                #     newfile2 =f.cleaned_data['reconocimientoacademico1']

                if DetalleTitulacionBachiller.objects.filter(titulacion=titulacion).exists():
                    detalle = DetalleTitulacionBachiller.objects.filter(titulacion=titulacion)[0]
                    detalle.titulacion = titulacion
                    detalle.calificacion = f.cleaned_data['calificacion']
                    detalle.actagrado = newfile if newfile else detalle.actagrado
                    detalle.anioinicioperiodograduacion = f.cleaned_data['anioinicioperiodograduacion']
                    detalle.aniofinperiodograduacion = f.cleaned_data['aniofinperiodograduacion']
                    detalle.reconocimientoacademico = newfile2 if newfile2 else detalle.reconocimientoacademico
                    detalle.save(request)
                    titulacion.registroarchivo = newfile if newfile else detalle.actagrado
                    titulacion.save(request)

                else:
                    detalle = DetalleTitulacionBachiller(titulacion=titulacion,
                                                         calificacion=f.cleaned_data['calificacion'],
                                                         actagrado=newfile,
                                                         anioinicioperiodograduacion=f.cleaned_data[
                                                             'anioinicioperiodograduacion'],
                                                         aniofinperiodograduacion=f.cleaned_data[
                                                             'aniofinperiodograduacion'],
                                                         reconocimientoacademico=newfile2)
                    titulacion.registroarchivo = newfile
                    titulacion.save(request)

                    detalle.save(request)

                # para tambien porder registrar en archivo
                archivobachiller = Archivo.objects.filter(tipo_id=16, inscripcion__persona=persona, status=True)
                archivoreconocimiento = Archivo.objects.filter(tipo_id=18, inscripcion__persona=persona,
                                                               status=True)
                inscri = perfilprincipal.inscripcion if perfilprincipal.es_estudiante() else None
                profe = perfilprincipal.inscripcion if perfilprincipal.es_profesor() else None

                if not archivobachiller and newfile:  # si no exsite en Archivo, lo guarda en tabla archivo sino lo actualiza
                    nombrearchivo = f'ACTA DE GRADO DE BACHILLER DE LA PERSONA: {persona}'

                    archivobachi = Archivo(
                        tipo_id=16,
                        nombre=nombrearchivo,
                        fecha=datetime.now().date(),
                        archivo=newfile,
                        aprobado=True,
                        inscripcion=inscri,
                        profesor=profe,
                        persona=persona,
                        sga=True
                    )
                    archivobachi.save(request)
                else:
                    if newfile:
                        archivobachi = archivobachiller.last()
                        archivobachi.inscripcion = inscri
                        archivobachi.archivo = newfile
                        archivobachi.save(request)

                if not archivoreconocimiento and newfile2:  # si no exsite en Archivo, lo guarda en tabla archivo sino lo actualiza
                    nombrearchivo = f'RECONOCIMIENTO ACADÉMICO DE LA PERSONA: {persona}'

                    archivorecono = Archivo(
                        tipo_id=18,
                        nombre=nombrearchivo,
                        fecha=datetime.now().date(),
                        archivo=newfile2,
                        aprobado=True,
                        inscripcion=inscri,
                        profesor=profe,
                        persona=persona,
                        sga=True
                    )
                    archivorecono.save(request)
                else:
                    if newfile2:
                        archivorecono = archivoreconocimiento.last()
                        archivorecono.inscripcion = inscri
                        archivorecono.archivo = newfile2
                        archivorecono.save(request)

                messages.success(request, 'Se guardó exitosamente.')
                log(u'Adiciono titulacion Bachiller: %s %s %s %s' % (
                    detalle, detalle.calificacion, detalle.anioinicioperiodograduacion,
                    detalle.aniofinperiodograduacion), request, "add")
                return JsonResponse({'result': False, 'mensaje': 'Guardado con exito'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': f'Error al guardar los datos {ex}'})

        # PROYECTOS

        elif action == 'addproyectoexterno':
            try:
                f = ProyectoInvestigacionExternoForm(request.POST, request.FILES)

                if not f.is_valid():
                    for k, v in f.errors.items():
                        raise NameError(k + ', ' + v[0])

                # Validación de campos
                if not 'archivo' in request.FILES:
                    raise NameError('Debe elegir el archivo del proyecto en formato PDF')

                if f.cleaned_data['fechainicio'] >= f.cleaned_data['fechafin']:
                    return JsonResponse({"result": "bad", "mensaje": u"Atención!!! La fecha de fin deber ser mayor a la fecha de inicio", "typewarning": True})

                # Verificar que el proyecto no esté repetido
                if persona.proyectoinvestigacionexterno_set.filter(nombre__iexact=f.cleaned_data['nombre'].strip()).exists():
                    return JsonResponse({"result": "bad", "mensaje": u"Atención!!! El proyecto de investigación ya se encuentra registrado", "typewarning": True})

                newfile = request.FILES['archivo']
                newfile._name = generar_nombre("proyinves_", newfile._name)

                # Guardo el proyecto
                proyectoexterno = ProyectoInvestigacionExterno(
                    persona=persona,
                    nombre=f.cleaned_data['nombre'],
                    rol=f.cleaned_data['rol'],
                    institucion=f.cleaned_data['institucion'],
                    financiamiento=f.cleaned_data['financiamiento'],
                    fechainicio=f.cleaned_data['fechainicio'],
                    fechafin=f.cleaned_data['fechafin'],
                    archivo=newfile
                )
                proyectoexterno.save(request)

                log(u'%s agregó proyecto de investigación externo: %s' % (persona, proyectoexterno), request, "add")

                return JsonResponse({"result": False}, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos. [%s]" % (ex)})

        elif action == 'editproyectoexterno':
            try:
                f = ProyectoInvestigacionExternoForm(request.POST, request.FILES)

                if not f.is_valid():
                    for k, v in f.errors.items():
                        raise NameError(k + ', ' + v[0])

                # Validación de campos
                if f.cleaned_data['fechainicio'] >= f.cleaned_data['fechafin']:
                    return JsonResponse({"result": "bad", "mensaje": u"Atención!!! La fecha de fin deber ser mayor a la fecha de inicio", "typewarning": True})

                # Verificar que el proyecto no esté repetido
                if persona.proyectoinvestigacionexterno_set.filter(nombre__iexact=f.cleaned_data['nombre'].strip()).exclude(pk=int(encrypt(request.POST['id']))).exists():
                    return JsonResponse({"result": "bad", "mensaje": u"Atención!!! El proyecto de investigación ya se encuentra registrado", "typewarning": True})
                
                # Consulto proyecto de investigación
                proyectoexterno = ProyectoInvestigacionExterno.objects.get(pk=int(encrypt(request.POST['id'])))

                # Actualizo proyecto externo
                proyectoexterno.nombre = f.cleaned_data['nombre']
                proyectoexterno.rol = f.cleaned_data['rol']
                proyectoexterno.institucion = f.cleaned_data['institucion']
                proyectoexterno.financiamiento = f.cleaned_data['financiamiento']
                proyectoexterno.fechainicio = f.cleaned_data['fechainicio']
                proyectoexterno.fechafin = f.cleaned_data['fechafin']

                if 'archivo' in request.FILES:
                    newfile = request.FILES['archivo']
                    newfile._name = generar_nombre("proyinves_", newfile._name)
                    proyectoexterno.archivo = newfile

                proyectoexterno.save(request)

                log(u'%s editó proyecto de investigación externo: %s' % (persona, proyectoexterno), request, "edit")

                return JsonResponse({"result": False}, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos. [%s]" % (ex)})

        elif action == 'delproyectoexterno':
            try:
                # Consulto proyecto de investigación
                proyectoexterno = ProyectoInvestigacionExterno.objects.get(pk=int(encrypt(request.POST['id'])))

                # Elimino el proyecto externo
                proyectoexterno.status = False
                proyectoexterno.save(request)

                log(u'%s eliminó proyecto de investigación externo: %s' % (persona, proyectoexterno), request, "del")
                res_js = {'error': False}
            except Exception as ex:
                transaction.set_rollback(True)
                err = 'Error on line {}'.format(sys.exc_info()[-1].tb_lineno)
                msg_err = f"Ocurrio un error: {ex.__str__()}. {err}"
                res_js = {'error': True, 'message': msg_err}
            return JsonResponse(res_js)

        # EXPERIENCIA

        elif action == 'addexperiencia':
            try:
                persona = request.session['persona']
                form = ExperienciaLaboralForm(request.POST, request.FILES)
                if 'archivo' in request.FILES:
                    d = request.FILES['archivo']
                    if d.size > 4194304:
                        return JsonResponse({"result": "bad", "mensaje": u"Error, archivo mayor a 4 Mb."})
                    else:
                        newfiles = request.FILES['archivo']
                        newfilesd = newfiles._name
                        ext = newfilesd[newfilesd.rfind("."):]
                        if not (ext == '.pdf' or ext == '.png' or ext == '.jpg' or ext == '.jpeg'):
                            return JsonResponse(
                                {"result": "bad", "mensaje": u"Error, solo archivos .pdf, png, jpg, jpeg."})
                if not form.is_valid():
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()],
                                         "mensaje": "Error en el formulario"})
                motivosalida = None
                fechafin = None
                if not form.cleaned_data['vigente']:
                    motivosalida = form.cleaned_data['motivosalida']
                    fechafin = form.cleaned_data['fechafin']
                experiencialaboral = ExperienciaLaboral(persona=persona,
                                                        tipoinstitucion=form.cleaned_data['tipoinstitucion'],
                                                        institucion=form.cleaned_data['institucion'],
                                                        cargo=form.cleaned_data['cargo'],
                                                        departamento=form.cleaned_data['departamento'],
                                                        pais=form.cleaned_data['pais'],
                                                        provincia=form.cleaned_data['provincia'],
                                                        canton=form.cleaned_data['canton'],
                                                        parroquia=form.cleaned_data['parroquia'],
                                                        fechainicio=form.cleaned_data['fechainicio'],
                                                        fechafin=fechafin,
                                                        motivosalida=motivosalida,
                                                        regimenlaboral=form.cleaned_data['regimenlaboral'],
                                                        horassemanales=form.cleaned_data['horassemanales'],
                                                        dedicacionlaboral=form.cleaned_data['dedicacionlaboral'],
                                                        actividadlaboral=form.cleaned_data['actividadlaboral'],
                                                        observaciones=form.cleaned_data['observaciones'])
                experiencialaboral.save(request)
                if 'archivo' in request.FILES:
                    newfile = request.FILES['archivo']
                    newfile._name = generar_nombre("experiencialaboral_", newfile._name)
                    experiencialaboral.archivo = newfile
                    experiencialaboral.save(request)
                    if DistributivoPersona.objects.filter(status=True, persona=persona):
                        lista = ['talento_humano@unemi.edu.ec', 'portizg@unemi.edu.ec']
                        asunto = "Ingresaron nueva experiencia laboral (archivo)"
                        send_html_mail(asunto, "emails/ingreso_archivos_tthh.html",
                                       {'asunto': asunto, 'd': experiencialaboral.persona.nombre_completo_inverso(),
                                        'fecha': datetime.now().date(), 'escenario': 'experiencia laboral'}, lista,
                                       [],
                                       cuenta=CUENTAS_CORREOS[1][1])
                log(u'Adiciono experiencia laboral: %s' % experiencialaboral, request, "add")
                return JsonResponse({'result': False, 'mensaje': 'Guardado con exito'})
            except Exception as ex:
                msg = ex.__str__()
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': f'Error al guardar los datos {ex}'})

        elif action == 'editexperiencia':
            try:
                persona = request.session['persona']
                form = ExperienciaLaboralForm(request.POST, request.FILES)
                if 'archivo' in request.FILES:
                    d = request.FILES['archivo']
                    if d.size > 4194304:
                        return JsonResponse({"result": "bad", "mensaje": u"Error, archivo mayor a 4 Mb."})
                    else:
                        newfiles = request.FILES['archivo']
                        newfilesd = newfiles._name
                        ext = newfilesd[newfilesd.rfind("."):]
                        if not (ext == '.pdf' or ext == '.png' or ext == '.jpg' or ext == '.jpeg'):
                            return JsonResponse(
                                {"result": "bad", "mensaje": u"Error, solo archivos .pdf, png, jpg, jpeg."})
                if not form.is_valid():
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()],
                                         "mensaje": "Error en el formulario"})
                motivosalida = None
                fechafin = None
                if not form.cleaned_data['vigente']:
                    motivosalida = form.cleaned_data['motivosalida']
                    fechafin = form.cleaned_data['fechafin']
                experiencialaboral = ExperienciaLaboral.objects.get(pk=encrypt_id(request.POST['id']))
                experiencialaboral.tipoinstitucion = form.cleaned_data['tipoinstitucion']
                experiencialaboral.institucion = form.cleaned_data['institucion']
                experiencialaboral.cargo = form.cleaned_data['cargo']
                experiencialaboral.departamento = form.cleaned_data['departamento']
                experiencialaboral.pais = form.cleaned_data['pais']
                experiencialaboral.provincia = form.cleaned_data['provincia']
                experiencialaboral.canton = form.cleaned_data['canton']
                experiencialaboral.parroquia = form.cleaned_data['parroquia']
                experiencialaboral.fechainicio = form.cleaned_data['fechainicio']
                experiencialaboral.fechafin = fechafin
                experiencialaboral.motivosalida = motivosalida
                experiencialaboral.regimenlaboral = form.cleaned_data['regimenlaboral']
                experiencialaboral.horassemanales = form.cleaned_data['horassemanales']
                experiencialaboral.dedicacionlaboral = form.cleaned_data['dedicacionlaboral']
                experiencialaboral.actividadlaboral = form.cleaned_data['actividadlaboral']
                experiencialaboral.observaciones = form.cleaned_data['observaciones']
                experiencialaboral.save(request)
                if 'archivo' in request.FILES:
                    newfile = request.FILES['archivo']
                    newfile._name = generar_nombre("experiencialaboral_", newfile._name)
                    experiencialaboral.archivo = newfile
                    experiencialaboral.save(request)
                log(u'Modifico experiencia laboral: %s' % experiencialaboral, request, "edit")
                return JsonResponse({'result': False, 'mensaje': 'Guardado con exito'})
            except Exception as ex:
                msg = ex.__str__()
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': f'Error al guardar los datos {ex}'})

        elif action == 'delexperiencia':
            try:
                experiencia = ExperienciaLaboral.objects.get(pk=encrypt_id(request.POST['id']))
                log(u"Elimino experiencia laboral: %s" % experiencia, request, "del")
                experiencia.delete()
                res_js = {'error': False}
            except Exception as ex:
                transaction.set_rollback(True)
                err = 'Error on line {}'.format(sys.exc_info()[-1].tb_lineno)
                msg_err = f"Ocurrio un error: {ex.__str__()}. {err}"
                res_js = {'error': True, 'mensaje': msg_err}
            return JsonResponse(res_js)

        # MERITOS

        elif action == 'addmerito':
            try:
                form = OtroMeritoForm(request.POST, request.FILES)
                if not form.is_valid():
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()],
                                         "mensaje": "Error en el formulario"})
                newfile = request.FILES['archivo']
                if newfile.size > 4194304:
                    return JsonResponse(
                        {"result": "bad", "mensaje": u"Error, Tamaño de archivo Maximo permitido es de 4Mb"})
                newfile._name = generar_nombre("documentos_", newfile._name)
                omerito = OtroMerito(persona=persona, nombre=form.cleaned_data['nombre'],
                                     fecha=form.cleaned_data['fecha'], institucion=form.cleaned_data['institucion'],
                                     archivo=newfile)
                omerito.save(request)
                log(u'Adicionó un nuevo merito a la hoja de vida: %s - la persona: %s' % (omerito, persona),request, "add")
                return JsonResponse({'result': False, 'mensaje': 'Guardado con exito'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': f'Error al guardar los datos {ex}'})

        elif action == 'editmerito':
            try:
                form = OtroMeritoForm(request.POST, request.FILES)
                if not form.is_valid():
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()],
                                         "mensaje": "Error en el formulario"})
                newfile = None
                omerito = OtroMerito.objects.get(pk=encrypt_id(request.POST['id']))
                if 'archivo' in request.FILES:
                    newfile = request.FILES['archivo']
                    if newfile.size > 4194304:
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"Error, Tamaño de archivo Maximo permitido es de 4Mb"})
                    newfile._name = generar_nombre("documentos_", newfile._name)
                omerito.nombre = form.cleaned_data['nombre']
                omerito.fecha = form.cleaned_data['fecha']
                omerito.institucion = form.cleaned_data['institucion']
                if newfile:
                    omerito.archivo = newfile
                omerito.save(request)
                log(u'Editó un merito a la hoja de vida: %s - la persona: %s' % (omerito, persona), request, "edit")
                return JsonResponse({'result': False, 'mensaje': 'Guardado con exito'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': f'Error al guardar los datos {ex}'})

        elif action == 'delmerito':
            try:
                omerito = OtroMerito.objects.get(pk=encrypt_id(request.POST['id']))
                omerito.status=False
                omerito.save(request)
                log(u'Elimino merito a la hoja de vida: %s - la persona: %s' % (omerito, persona), request, "del")
                res_js = {'error': False}
            except Exception as ex:
                transaction.set_rollback(True)
                err = 'Error on line {}'.format(sys.exc_info()[-1].tb_lineno)
                msg_err = f"Ocurrio un error: {ex.__str__()}. {err}"
                res_js = {'error': True, 'mensaje': msg_err}
            return JsonResponse(res_js)

        # TUTORIAS

        elif action == 'addcertificadotutoria':
            try:
                persona = request.session['persona']
                f = CertificadoTutoriaForm(request.POST, request.FILES)
                if 'archivo' in request.FILES:
                    d = request.FILES['archivo']
                    newfilesd = d._name
                    ext = newfilesd[newfilesd.rfind("."):]
                    if ext == '.pdf':
                        a = 1
                    else:
                        return JsonResponse({"result": "bad", "mensaje": u"Error, solo archivos .pdf."})
                    if d.size > 2194304:
                        return JsonResponse({"result": "bad", "mensaje": u"Error, archivo mayor a 2 Mb."})
                if not f.is_valid():
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()],
                                         "mensaje": "Error en el formulario"})
                if not (f.cleaned_data['fechainicio'] <= f.cleaned_data['fechafin']):
                    raise NameError("Las fechas no son validas.")

                certificado = CertificadoTutoriaHV(persona=persona,
                                                   institucion=f.cleaned_data['institucion'],
                                                   nombreproyecto=f.cleaned_data['nombreproyecto'],
                                                   descripcion=f.cleaned_data['descripcion'],
                                                   calificacion=f.cleaned_data['calificacion'],
                                                   fechainicio=f.cleaned_data['fechainicio'],
                                                   fechafin=f.cleaned_data['fechafin'])
                certificado.save(request)
                newfile = None
                if 'archivo' in request.FILES:
                    newfile = request.FILES['archivo']
                    newfile._name = generar_nombre("certificado_tutoria_", newfile._name)
                    certificado.archivo = newfile
                    certificado.save(request)
                log(u'Adicionó certificado de tutoria: %s' % persona, request, "add")
                return JsonResponse({'result': False, 'mensaje': 'Guardado con exito'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': f'Error al guardar los datos {ex}'})

        elif action == 'editcertificadotutoria':
            try:
                persona = request.session['persona']
                certificado = CertificadoTutoriaHV.objects.get(pk=encrypt_id(request.POST['id']))
                if 'archivo' in request.FILES:
                    d = request.FILES['archivo']
                    if d.size > 2194304:
                        return JsonResponse({"result": True, "mensaje": u"Error, archivo mayor a 2 Mb."})
                    else:
                        newfiles = request.FILES['archivo']
                        newfilesd = newfiles._name
                        ext = newfilesd[newfilesd.rfind("."):]
                        if not (ext == '.pdf'):
                            return JsonResponse({"result": "bad", "mensaje": u"Error, solo archivos .pdf"})
                f = CertificadoTutoriaForm(request.POST, request.FILES)
                if not f.is_valid():
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in f.errors.items()],
                                         "mensaje": "Error en el formulario"})
                if not (f.cleaned_data['fechainicio'] <= f.cleaned_data['fechafin']):
                    return JsonResponse({"result": True, "mensaje": u"Las fechas no son validas."})
                certificado.institucion = f.cleaned_data['institucion']
                certificado.nombreproyecto = f.cleaned_data['nombreproyecto']
                certificado.descripcion = f.cleaned_data['descripcion']
                certificado.fechainicio = f.cleaned_data['fechainicio']
                certificado.fechafin = f.cleaned_data['fechafin']
                certificado.calificacion = f.cleaned_data['calificacion']
                if 'archivo' in request.FILES:
                    newfile = request.FILES['archivo']
                    newfile._name = generar_nombre("certificado_tutoria_", newfile._name)
                    certificado.archivo = newfile
                certificado.save(request)
                log(u'Modifico certificación de tutoria personal: %s' % persona, request, "edit")
                return JsonResponse({'result': False, 'mensaje': 'Guardado con exito'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': f'Error al guardar los datos {ex}'})

        elif action == 'delcertificadotutoria':
            try:
                certificadopersona = CertificadoTutoriaHV.objects.get(pk=encrypt_id(request.POST['id']))
                certificadopersona.status = False
                certificadopersona.save(request)
                log(u'Eliminó un certificado de tutoria en hoja de vida: %s - la persona: %s' % (certificadopersona, persona), request, "add")
                res_js = {'error': False}
            except Exception as ex:
                transaction.set_rollback(True)
                err = 'Error on line {}'.format(sys.exc_info()[-1].tb_lineno)
                msg_err = f"Ocurrio un error: {ex.__str__()}. {err}"
                res_js = {'error': True, 'mensaje': msg_err}
            return JsonResponse(res_js)

        # BECAS

        elif action == 'addbecaexterna':
            try:
                if 'archivo' in request.FILES:
                    arch = request.FILES['archivo']
                    extension = arch._name.split('.')
                    tam = len(extension)
                    exte = extension[tam - 1]
                    if arch.size > 4194304:
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"Error, el tamaño del archivo es mayor a 4 Mb."})
                    if not exte.lower() in ['pdf']:
                        return JsonResponse({"result": "bad", "mensaje": u"Solo se permiten archivos .pdf"})

                form = DatosBecaForm(request.POST, request.FILES)
                if not form.is_valid():
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()],
                                     "mensaje": "Error en el formulario"})
                beca = BecaPersona(persona=persona,
                                   tipoinstitucion=form.cleaned_data['tipoinstitucion'],
                                   institucion=form.cleaned_data['institucion'],
                                   fechainicio=form.cleaned_data['fechainicio'],
                                   fechafin=form.cleaned_data['fechafin']
                                   )
                beca.save(request)

                if 'archivo' in request.FILES:
                    arch._name = generar_nombre("archivobeca", arch._name)
                    beca.archivo = arch
                    beca.estadoarchivo = 1
                    beca.save(request)
                log(u'Agregó datos de la beca: %s - la persona: %s' % (beca, persona), request, "add")
                return JsonResponse({'result': False, 'mensaje': 'Guardado con exito'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': f'Error al guardar los datos {ex}'})

        elif action == 'editbecaexterna':
            try:
                if 'archivo' in request.FILES:
                    arch = request.FILES['archivo']
                    extension = arch._name.split('.')
                    tam = len(extension)
                    exte = extension[tam - 1]
                    if arch.size > 4194304:
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"Error, el tamaño del archivo es mayor a 4 Mb."})
                    if not exte.lower() in ['pdf']:
                        return JsonResponse({"result": "bad", "mensaje": u"Solo se permiten archivos .pdf"})

                form = DatosBecaForm(request.POST, request.FILES)
                if not form.is_valid():
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()],
                                     "mensaje": "Error en el formulario"})
                beca = BecaPersona.objects.get(pk=encrypt_id(request.POST['id']))
                if beca.fechainicio >= form.cleaned_data['fechafin']:
                    return JsonResponse({"result": "bad","mensaje": u"La fecha de fin de la beca debe ser mayor a la fecha rige"})
                beca.tipoinstitucion = form.cleaned_data['tipoinstitucion']
                beca.institucion = form.cleaned_data['institucion']
                beca.fechainicio = form.cleaned_data['fechainicio']
                beca.estadoarchivo = 1
                beca.fechafin = form.cleaned_data['fechafin']
                if 'archivo' in request.FILES:
                    arch._name = generar_nombre("archivobeca", arch._name)
                    beca.archivo = arch
                beca.save(request)
                log(u'Editó datos de la beca: %s - la persona: %s' % (beca, persona), request, "edit")
                return JsonResponse({'result': False, 'mensaje': 'Guardado con exito'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': f'Error al guardar los datos {ex}'})

        elif action == 'delbecaexterna':
            try:
                beca = BecaPersona.objects.get(pk=request.POST['id'])
                beca.status = False
                beca.save(request)
                log(u'Elimino beca externa: %s' % beca, request, "del")
                res_js = {'error': False}
            except Exception as ex:
                transaction.set_rollback(True)
                err = 'Error on line {}'.format(sys.exc_info()[-1].tb_lineno)
                msg_err = f"Ocurrio un error: {ex.__str__()}. {err}"
                res_js = {'error': True, 'mensaje': msg_err}
            return JsonResponse(res_js)

        # INFORMES MENSAULES

        elif action == 'addinformemensual':
            try:
                form = InformeMensualForm(request.POST, request.FILES)
                arch = request.FILES['archivo']
                extencion = arch._name.split('.')
                exte = extencion[1]
                if arch.size > 10485760:
                    return JsonResponse({"result": True, "mensaje": u"Error, archivo mayor a 10 Mb."})
                if not exte == 'pdf' or not exte == 'PDF':
                    a = 1
                else:
                    return JsonResponse({"result": True, "mensaje": u"Solo archivos .pdf"})
                if not form.is_valid():
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()],
                                         "mensaje": "Error en el formulario"})
                # nfile = request.FILES['archivo']
                # nfile._name = generar_nombre("documentos_", nfile._name)
                newfile = request.FILES['archivo']
                newfile._name = generar_nombre("informe_", newfile._name)
                informe = InformeMensual(persona=persona,
                                         fechainicio=form.cleaned_data['fechainicio'],
                                         fechafin=form.cleaned_data['fechafin'],
                                         observacion=form.cleaned_data['observacion'],
                                         archivo=newfile)
                informe.save(request)
                log(u'Adiciono un inform mensual: %s' % informe, request, "add")
                return JsonResponse({'result': False, 'mensaje': 'Guardado con exito'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': f'Error al guardar los datos {ex}'})

        elif action == 'editinformemensual':
            try:
                form = InformeMensualForm(request.POST, request.FILES)
                if 'archivo' in request.FILES:
                    arch = request.FILES['archivo']
                    extencion = arch._name.split('.')
                    exte = extencion[1]
                    if arch.size > 10485760:
                        return JsonResponse({"result": "bad", "mensaje": u"Error, archivo mayor a 10 Mb."})
                    if not exte == 'pdf':
                        return JsonResponse({"result": "bad", "mensaje": u"Solo archivos .pdf"})
                if not form.is_valid():
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()],
                                         "mensaje": "Error en el formulario"})
                informe = persona.informemensual_set.get(pk=encrypt_id(request.POST['id']))
                informe.fechainicio = form.cleaned_data['fechainicio']
                informe.fechafin = form.cleaned_data['fechafin']
                informe.observacion = form.cleaned_data['observacion']
                if 'archivo' in request.FILES:
                    newfile = request.FILES['archivo']
                    newfile._name = generar_nombre("informe_", newfile._name)
                    informe.archivo = newfile
                informe.save(request)
                log(u'Editó un inform mensual: %s' % informe, request, "add")
                return JsonResponse({'result': False, 'mensaje': 'Guardado con exito'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': f'Error al guardar los datos {ex}'})

        elif action == 'delinformemensual':
            try:
                informe = InformeMensual.objects.get(pk=encrypt_id(request.POST['id']))
                informe.status=True
                informe.save(request)
                log(u'Eliminó un inform mensual: %s' % informe, request, "add")
                res_js = {'error': False}
            except Exception as ex:
                transaction.set_rollback(True)
                err = 'Error on line {}'.format(sys.exc_info()[-1].tb_lineno)
                msg_err = f"Ocurrio un error: {ex.__str__()}. {err}"
                res_js = {'error': True, 'mensaje': msg_err}
            return JsonResponse(res_js)

        # DATOS DEPORTISTA

        elif action == 'deldatosdeportista':
            try:
                deportista = DeportistaPersona.objects.get(pk=encrypt_id(request.POST['id']))
                deportista.status = False
                deportista.save(request)
                log(u'Elimino deportista alto rendimiento: %s' % deportista, request, "del")
                res_js = {'error': False}
            except Exception as ex:
                transaction.set_rollback(True)
                err = 'Error on line {}'.format(sys.exc_info()[-1].tb_lineno)
                msg_err = f"Ocurrio un error: {ex.__str__()}. {err}"
                res_js = {'error': True, 'mensaje': msg_err}
            return JsonResponse(res_js)

        elif action == 'adddatosdeportista':
            try:
                if 'archivoevento' in request.FILES:
                    arch = request.FILES['archivoevento']
                    extension = arch._name.split('.')
                    tam = len(extension)
                    exte = extension[tam - 1]
                    if arch.size > 4194304:
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"Error, el tamaño del archivo del evento es mayor a 4 Mb."})
                    if not exte.lower() in ['pdf']:
                        return JsonResponse({"result": "bad", "mensaje": u"Solo se permiten archivos .pdf"})

                if 'archivoentrena' in request.FILES:
                    arch2 = request.FILES['archivoentrena']
                    extension = arch2._name.split('.')
                    tam = len(extension)
                    exte = extension[tam - 1]
                    if arch2.size > 4194304:
                        return JsonResponse({"result": "bad",
                                             "mensaje": u"Error, el tamaño del archivo del entrenamiento es mayor a 4 Mb."})
                    if not exte.lower() in ['pdf']:
                        return JsonResponse({"result": "bad", "mensaje": u"Solo se permiten archivos .pdf"})

                form = DeportistaForm(request.POST, request.FILES)

                if not form.is_valid():
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()],
                                         "mensaje": "Error en el formulario"})

                if form.cleaned_data['fechainicioevento'] > form.cleaned_data['fechafinevento']:
                    return JsonResponse({"result": "bad",
                                         "mensaje": u"La fecha de inicio de evento debe ser menor o igual a la fecha de fin"})

                if form.cleaned_data['fechainicioentrena'] > form.cleaned_data['fechafinentrena']:
                    return JsonResponse({"result": "bad",
                                         "mensaje": u"La fecha de inicio de entrenamiento debe ser menor o igual a la fecha de fin"})

                deportista = DeportistaPersona(persona=persona,
                                               representapais=form.cleaned_data['representapais'],
                                               evento=form.cleaned_data['evento'],
                                               paisevento=form.cleaned_data['paisevento'],
                                               equiporepresenta=form.cleaned_data['equiporepresenta'],
                                               fechainicioevento=form.cleaned_data['fechainicioevento'],
                                               fechafinevento=form.cleaned_data['fechafinevento'],
                                               fechainicioentrena=form.cleaned_data['fechainicioentrena'],
                                               fechafinentrena=form.cleaned_data['fechafinentrena'],
                                               vigente=1)
                deportista.save(request)

                if 'archivoevento' in request.FILES:
                    arch._name = generar_nombre("archivodepevento", arch._name)
                    deportista.archivoevento = arch
                    deportista.estadoarchivoevento = 1

                if 'archivoentrena' in request.FILES:
                    arch2._name = generar_nombre("archivodepentrena", arch2._name)
                    deportista.archivoentrena = arch2
                    deportista.estadoarchivoentrena = 1

                deportista.save(request)
                deportista.disciplina.clear()
                for dis in form.cleaned_data['disciplina']:
                    deportista.disciplina.add(dis)
                deportista.save(request)
                log(u'Agregó datos de deportista: %s - la persona: %s' % (deportista, persona), request, "add")
                return JsonResponse({'result': False, 'mensaje': 'Guardado con exito'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': f'Error al guardar los datos {ex}'})

        elif action == 'editdatosdeportista':
            try:
                if 'archivoevento' in request.FILES:
                    arch = request.FILES['archivoevento']
                    extension = arch._name.split('.')
                    tam = len(extension)
                    exte = extension[tam - 1]
                    if arch.size > 4194304:
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"Error, el tamaño del archivo del evento es mayor a 4 Mb."})
                    if not exte.lower() in ['pdf']:
                        return JsonResponse({"result": "bad", "mensaje": u"Solo se permiten archivos .pdf"})

                if 'archivoentrena' in request.FILES:
                    arch2 = request.FILES['archivoentrena']
                    extension = arch2._name.split('.')
                    tam = len(extension)
                    exte = extension[tam - 1]
                    if arch2.size > 4194304:
                        return JsonResponse({"result": "bad",
                                             "mensaje": u"Error, el tamaño del archivo del entrenamiento es mayor a 4 Mb."})
                    if not exte.lower() in ['pdf']:
                        return JsonResponse({"result": "bad", "mensaje": u"Solo se permiten archivos .pdf"})

                form = DeportistaForm(request.POST, request.FILES)
                if not form.is_valid():
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()],
                                         "mensaje": "Error en el formulario"})

                deportista = DeportistaPersona.objects.get(pk=encrypt_id(request.POST['id']))
                if form.cleaned_data['representapais']:
                    if form.cleaned_data['fechainicioevento'] > form.cleaned_data['fechafinevento']:
                        return JsonResponse({"result": "bad",
                                             "mensaje": u"La fecha de inicio de evento debe ser menor o igual a la fecha de fin"})

                    if form.cleaned_data['fechainicioentrena'] > form.cleaned_data['fechafinentrena']:
                        return JsonResponse({"result": "bad",
                                             "mensaje": u"La fecha de inicio de entrenamiento debe ser menor o igual a la fecha de fin"})

                    deportista.representapais = form.cleaned_data['representapais']
                    deportista.evento = form.cleaned_data['evento']
                    deportista.paisevento = form.cleaned_data['paisevento']
                    deportista.equiporepresenta = form.cleaned_data['equiporepresenta']
                    deportista.fechainicioevento = form.cleaned_data['fechainicioevento']
                    deportista.fechafinevento = form.cleaned_data['fechafinevento']
                    deportista.fechainicioentrena = form.cleaned_data['fechainicioentrena']
                    deportista.fechafinentrena = form.cleaned_data['fechafinentrena']

                deportista.vigente = form.cleaned_data['vigente']

                if 'archivoevento' in request.FILES:
                    arch._name = generar_nombre("archivodepevento", arch._name)
                    deportista.archivoevento = arch
                    deportista.estadoarchivoevento = 1

                if 'archivoentrena' in request.FILES:
                    arch2._name = generar_nombre("archivodepentrena", arch2._name)
                    deportista.archivoentrena = arch2
                    deportista.estadoarchivoentrena = 1

                deportista.disciplina.clear()
                for dis in form.cleaned_data['disciplina']:
                    deportista.disciplina.add(dis)
                deportista.save(request)

                log(u'Editó datos de deportista: %s - la persona: %s' % (deportista, persona), request, "edit")
                return JsonResponse({"result": False})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": f"Error al guardar los datos: {ex}"})

        # DATOS ARTISTA

        elif action == 'adddatosartista':
            try:
                if 'archivo' in request.FILES:
                    arch = request.FILES['archivo']
                    extension = arch._name.split('.')
                    tam = len(extension)
                    exte = extension[tam - 1]
                    if arch.size > 4194304:
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"Error, el tamaño del archivo es mayor a 4 Mb."})
                    if not exte.lower() in ['pdf']:
                        return JsonResponse({"result": "bad", "mensaje": u"Solo se permiten archivos .pdf"})

                form = ArtistaForm(request.POST, request.FILES)
                if not form.is_valid():
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()],
                                         "mensaje": "Error en el formulario"})
                if form.cleaned_data['fechainicioensayo']:
                    if form.cleaned_data['fechainicioensayo'] > form.cleaned_data['fechafinensayo']:
                        return JsonResponse({"result": "bad",
                                             "mensaje": u"La fecha de inicio de ensayo debe ser menor o igual a la fecha de fin"})

                artista = ArtistaPersona(persona=persona,
                                         grupopertenece=form.cleaned_data['grupopertenece'],
                                         fechainicioensayo=form.cleaned_data['fechainicioensayo'],
                                         fechafinensayo=form.cleaned_data['fechafinensayo'],
                                         vigente=1
                                         )
                artista.save(request)

                if 'archivo' in request.FILES:
                    arch._name = generar_nombre("archivoartista", arch._name)
                    artista.archivo = arch
                    artista.estadoarchivo = 1

                artista.save(request)
                for campo in form.cleaned_data['campoartistico']:
                    artista.campoartistico.add(campo)
                artista.save(request)
                log(u'Agregó datos de artista: %s - la persona: %s' % (artista, persona), request, "add")

                return JsonResponse({'result': False, 'mensaje': 'Guardado con exito'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': f'Error al guardar los datos {ex}'})

        elif action == 'editdatosartista':
            try:
                if 'archivo' in request.FILES:
                    arch = request.FILES['archivo']
                    extension = arch._name.split('.')
                    tam = len(extension)
                    exte = extension[tam - 1]
                    if arch.size > 4194304:
                        return JsonResponse(
                            {"result": "bad", "mensaje": u"Error, el tamaño del archivo es mayor a 4 Mb."})
                    if not exte.lower() in ['pdf']:
                        return JsonResponse({"result": "bad", "mensaje": u"Solo se permiten archivos .pdf"})

                form = ArtistaForm(request.POST, request.FILES)
                if not form.is_valid():
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()],
                                         "mensaje": "Error en el formulario"})
                artista = ArtistaPersona.objects.get(pk=int(encrypt(request.POST['id'])))
                if form.cleaned_data['grupopertenece']:
                    if form.cleaned_data['fechainicioensayo']:
                        if form.cleaned_data['fechainicioensayo'] > form.cleaned_data['fechafinensayo']:
                            return JsonResponse({"result": "bad",
                                                 "mensaje": u"La fecha de inicio de ensayo debe ser menor o igual a la fecha de fin"})

                    artista.grupopertenece = form.cleaned_data['grupopertenece']
                    artista.fechainicioensayo = form.cleaned_data['fechainicioensayo']
                    artista.fechafinensayo = form.cleaned_data['fechafinensayo']

                    if 'archivo' in request.FILES:
                        arch._name = generar_nombre("archivoartista", arch._name)
                        artista.archivo = arch
                        artista.estadoarchivo = 1

                artista.vigente = form.cleaned_data['vigente']
                artista.save(request)

                if form.cleaned_data['grupopertenece']:
                    artista.campoartistico.clear()
                    for campo in form.cleaned_data['campoartistico']:
                        artista.campoartistico.add(campo)
                    artista.save(request)

                log(u'Editó datos de artista: %s - la persona: %s' % (artista, persona), request, "edit")
                return JsonResponse({'result': False, 'mensaje': 'Guardado con exito'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': f'Error al guardar los datos {ex}'})

        elif action == 'deldatosartista':
            try:
                artista = ArtistaPersona.objects.get(pk=request.POST['id'])
                artista.status = False
                artista.save(request)
                log(u'Elimino datos de artista: %s' % artista, request, "del")
                res_js = {'error': False}
            except Exception as ex:
                transaction.set_rollback(True)
                err = 'Error on line {}'.format(sys.exc_info()[-1].tb_lineno)
                msg_err = f"Ocurrio un error: {ex.__str__()}. {err}"
                res_js = {'error': True, 'mensaje': msg_err}
            return JsonResponse(res_js)

        # ENFERMEDADES

        elif action == 'addenfermedad':
            with transaction.atomic():
                try:
                    form = PersonaEnfermedadForm(request.POST, request.FILES)
                    if not form.is_valid():
                        for k, v in form.errors.items():
                            raise NameError(v[0])
                        # raise NameError(u"Complete todos los campos vacios.")
                    if PersonaEnfermedad.objects.filter(persona=persona, enfermedad=form.cleaned_data['enfermedad'],
                                                        status=True).exists():
                        raise NameError(u"Enfermedad seleccionada ya se encuentra registrada.")
                    personaenfermedad = PersonaEnfermedad(persona=persona, enfermedad=form.cleaned_data['enfermedad'])
                    personaenfermedad.save(request)
                    if 'archivomedico' in request.FILES:
                        newfile = request.FILES['archivomedico']
                        extension = newfile._name.split('.')
                        tam = len(extension)
                        exte = extension[tam - 1]
                        if newfile.size > 2194304:
                            raise NameError(u"El tamaño del archivo es mayor a 2 Mb.")
                        if not exte.lower() in ['pdf']:
                            raise NameError(u"Solo archivos .pdf,.jpg, .jpeg")
                        newfile._name = generar_nombre(str(elimina_tildes(personaenfermedad.enfermedad)), newfile._name)
                        personaenfermedad.archivomedico = newfile
                        personaenfermedad.save(request)
                    log(u'Adiciono enfermedad: %s' % personaenfermedad, request, "addenfermedad")
                    return JsonResponse({"result": False}, safe=False)
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": True, "mensaje": f"Intentelo más tarde. {ex.__str__()}"}, safe=False)

        elif action == 'editenfermedad':
            try:
                with transaction.atomic():
                    personaenfermedad = PersonaEnfermedad.objects.get(pk=int(encrypt(request.POST['id'])))
                    form = PersonaEnfermedadForm(request.POST, request.FILES)
                    if form.is_valid():
                        if PersonaEnfermedad.objects.filter(persona=persona, enfermedad=form.cleaned_data['enfermedad'],
                                                            status=True).exclude(id=personaenfermedad.id).exists():
                            transaction.set_rollback(True)
                            return JsonResponse(
                                {"result": True, "mensaje": "Enfermedad seleccionada ya se encuentra registrada."},
                                safe=False)
                        personaenfermedad.enfermedad = form.cleaned_data['enfermedad']
                        personaenfermedad.save(request)
                        if 'archivomedico' in request.FILES:
                            newfile = request.FILES['archivomedico']
                            extension = newfile._name.split('.')
                            tam = len(extension)
                            exte = extension[tam - 1]
                            if newfile.size > 2194304:
                                transaction.set_rollback(True)
                                return JsonResponse(
                                    {"result": True, "mensaje": u"Error, el tamaño del archivo es mayor a 2 Mb."})
                            if not exte.lower() in ['pdf']:
                                transaction.set_rollback(True)
                                return JsonResponse(
                                    {"result": True, "mensaje": u"Error, solo archivos .pdf,.jpg, .jpeg"})
                            newfile._name = generar_nombre(str(personaenfermedad.enfermedad), newfile._name)
                            personaenfermedad.archivomedico = newfile
                            personaenfermedad.save(request)
                        log(u'Edicion de enfermedad de persona: %s' % personaenfermedad, request, "editenfermedad")
                        return JsonResponse({"result": False}, safe=False)
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({"result": True, "mensaje": "Complete todos los campos vacios."},
                                            safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        elif action == 'delenfermedad':
            try:
                with transaction.atomic():
                    enfermedad = PersonaEnfermedad.objects.get(pk=int(encrypt(request.POST['id'])))
                    enfermedad.status = False
                    enfermedad.save(request)
                    log(u'Elimino registro de enfermedad : %s - %s - %s', request, "delenfermedad")
                    res_json = {"error": False}
            except Exception as ex:
                res_json = {'error': True, "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)

        # SITUACION LABORAL

        elif action == 'datossituacionlaboral':
            try:
                persona = request.session['persona']
                if PersonaSituacionLaboral.objects.filter(persona_id=persona.id, status=True).exists():
                    situacionlaboral = PersonaSituacionLaboral.objects.get(persona_id=persona.id)
                else:
                    situacionlaboral = PersonaSituacionLaboral(persona=persona)
                    situacionlaboral.save(request)
                f = DatosSituacionLaboralForm(request.POST)
                if f.is_valid():
                    situacionlaboral.tipoinstitucionlaboral = 0
                    situacionlaboral.lugartrabajo = ''
                    situacionlaboral.negocio = ''
                    situacionlaboral.disponetrabajo = f.cleaned_data['disponetrabajo']
                    if situacionlaboral.disponetrabajo:
                        situacionlaboral.tipoinstitucionlaboral = f.cleaned_data['tipoinstitucionlaboral']
                        situacionlaboral.lugartrabajo = f.cleaned_data['lugartrabajo']
                    situacionlaboral.buscaempleo = f.cleaned_data['buscaempleo']
                    situacionlaboral.tienenegocio = f.cleaned_data['tienenegocio']
                    if situacionlaboral.tienenegocio:
                        situacionlaboral.negocio = f.cleaned_data['negocio']
                    situacionlaboral.save(request)

                    log(u'Modifico datos de situacion laboral: %s' % persona, request, "datossituacionlaboral")
                    return JsonResponse({'result': False, 'mensaje': 'Guardado con exito'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': f'{ex}'})

        elif action == 'saveAportacionHistorialLaboral':
            try:
                id = int(encrypt(request.POST['id'])) if 'id' in request.POST and request.POST['id'] and request.POST[
                    'id'] != '' else None
                f = PersonaAportacionHistorialLaboralForm(request.POST, request.FILES)
                typeForm = 'edit' if id else 'new'
                newfile = None
                if 'archivo_resumen' in request.FILES:
                    newfile = request.FILES['archivo_resumen']
                    if newfile:
                        newfilesd = newfile._name
                        ext = newfilesd[newfilesd.rfind("."):]
                        if not ext == '.pdf':
                            return JsonResponse({"result": "bad", "mensaje": u"Error, solo archivos .pdf."})
                        if newfile.size > 20480000:
                            return JsonResponse({"result": "bad", "mensaje": u"Error, archivo mayor a 20 Mb."})
                        if newfile:
                            newfile._name = generar_nombre("archivo_resumen", newfile._name)

                newfile2 = None
                if 'archivo_detalle' in request.FILES:
                    newfile2 = request.FILES['archivo_detalle']
                    if newfile2:
                        newfilesd2 = newfile2._name
                        ext2 = newfilesd2[newfilesd2.rfind("."):]
                        if not ext2 == '.pdf':
                            return JsonResponse({"result": "bad", "mensaje": u"Error, solo archivos .pdf."})
                        if newfile2.size > 20480000:
                            return JsonResponse({"result": "bad", "mensaje": u"Error, archivo mayor a 20 Mb."})
                        if newfile2:
                            newfile2._name = generar_nombre("archivo_detalle", newfile2._name)
                if not f.is_valid():
                    for k, v in f.errors.items():
                        raise NameError(v[0])
                registro = PersonaAportacionHistorialLaboral.objects.filter(pk=id).first()
                accion = 'edit'
                if not registro:
                    registro = PersonaAportacionHistorialLaboral()
                    accion = 'add'
                registro.persona = persona
                if newfile:
                    registro.archivo_resumen = newfile
                if newfile2:
                    registro.archivo_detalle = newfile2
                registro.save(request)
                log(u'%s nuevo Historial Laboral de Aportación: %s' % (
                'Adiciono' if accion == 'add' else 'Edito', registro), request, accion)
                return JsonResponse({"result": False, "mensaje": u"Registro guardado exitosamente."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": u"Error al guardar los datos. %s" % ex.__str__()})

        elif action == 'deleteAportacionHistorialLaboral':
            try:
                id = int(encrypt(request.POST['id'])) if 'id' in request.POST and request.POST['id'] and request.POST['id'] != '' else None
                if not id:
                    raise NameError('Parametro id no encontrado')
                registro = PersonaAportacionHistorialLaboral.objects.get(pk=id)
                registro.status = False
                registro.save(request)
                res_js = {'error': False}
            except Exception as ex:
                transaction.set_rollback(True)
                err = 'Error on line {}'.format(sys.exc_info()[-1].tb_lineno)
                msg_err = f"Ocurrio un error: {ex.__str__()}. {err}"
                res_js = {'error': True, 'mensaje': msg_err}
            return JsonResponse(res_js)

        elif action == 'firmainformemensualposgrado':
            try:
                idinforme = int(encrypt(request.POST['id_objeto']))
                hist_ = HistorialProcesoSolicitud.objects.get(status=True, id=idinforme)
                archivo_borrar = hist_.archivo.url
                certificado = request.FILES["firma"]
                contrasenaCertificado = request.POST['palabraclave']
                extension_certificado = os.path.splitext(certificado.name)[1][1:]
                bytes_certificado = certificado.read()
                texto = f'{persona}'
                x,y, numpaginafirma = obtener_posicion_x_y_saltolinea(hist_.archivo.url, texto)
                try:
                    datau = JavaFirmaEc(
                        archivo_a_firmar=hist_.archivo, archivo_certificado=bytes_certificado,
                        extension_certificado=extension_certificado,
                        password_certificado=contrasenaCertificado,
                        page=int(numpaginafirma), reason='', lx=x+50, ly=y+20
                    ).sign_and_get_content_bytes()
                except:
                    return JsonResponse({'result': True, 'mensaje': f'Ocurrio un error con FirmaEC'})
                documento_a_firmar = io.BytesIO()
                documento_a_firmar.write(datau)
                documento_a_firmar.seek(0)
                nombresinciales = ''
                nombre = persona.nombres.split()
                if len(nombre) > 1:
                    nombresiniciales = '{}{}'.format(nombre[0][0], nombre[1][0])
                else:
                    nombresiniciales = '{}'.format(nombre[0][0])
                inicialespersona = '{}{}{}'.format(nombresiniciales, persona.apellido1[0], persona.apellido2[0])
                name_file = f'informe_actividad_diaria_{inicialespersona.lower()}_{hist_.requisito.solicitud.numero}.pdf'
                hist_.archivo.save(f'{name_file.replace(".pdf", "")}_firmado.pdf',
                                   ContentFile(documento_a_firmar.read()))
                if os.path.exists(SITE_STORAGE + archivo_borrar):
                    os.remove(SITE_STORAGE + archivo_borrar)
                cuerpo = f"Informe mensual de posgrado generado y firmado por {hist_.requisito.solicitud.contrato.persona}"

                persona_notificacion = None
                redirect_mod = f'/adm_solicitudpago'

                requisito = hist_.requisito
                requisito.estado = 9#informe de actividad a estado firmado por profesional
                requisito.save(request)
                solicitud = requisito.solicitud
                contrato = solicitud.contrato
                if contrato.validadorgp:
                    persona_notificacion = contrato.validadorgp
                    redirect_mod = f'/adm_solicitudpago?action=viewinformesmen'
                    # estado_solicitud = 3
                else:
                    estado_solicitud = 6
                    persona_notificacion = contrato.gestion.responsable
                    solicitud.estado = estado_solicitud
                    solicitud.save(request)

                    noti("Solicitud de pago de %s para validación del informe mensual de posgrado" % hist_.requisito.solicitud.contrato.persona,cuerpo, persona_notificacion, None, redirect_mod,hist_.id,1, 'sga', hist_, request)
                    obshisto = HistorialObseracionSolicitudPago(
                        solicitud=solicitud,
                        observacion=cuerpo,
                        persona=persona,
                        estado=estado_solicitud,
                        fecha=datetime.now()
                    )
                    obshisto.save(request)
                    log(f'Firmo el informe mensual de la solicitud: {solicitud}', request, 'change')
                return JsonResponse({'result': False, 'to': f'/pro_solicitudpago?action=requisitos_solicitudes_pagos&id={encrypt(hist_.requisito.solicitud_id)}'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': f'Error de conexión. {ex.__str__()}'})

        elif action == 'informe-administrativo-posgrado':
            try:
                data['title'] = u"INFORME DE ACTIVIDADES ADMINISTRATIVAS"
                data['hoy'] = hoy = datetime.now()
                fi, ff = convertir_fecha_hora_invertida(request.POST['fi'] + " 00:01"), convertir_fecha_hora_invertida(request.POST['ff'] + " 23:59")
                fil_bitacora = Q(fecha__gte=fi, fecha__lte=ff, persona=persona, status=True)
                actividadesbitacora = BitacoraActividadDiaria.objects.filter(fil_bitacora).order_by('fecha')
                fil_contrato = Q(persona=persona, fechainicio__lte=fi, fechainicio__lt=ff, fechafin__gt=fi,fechafin__gte=ff, status=True)
                contrato = ContratoDip.objects.filter(fil_contrato).exclude(estado=5).order_by('-fecha_creacion').first()
                if contrato:
                    if not contrato.validadorgp:
                        raise NameError("No tiene asignado analista validador, contactar a la gestión de pagos.")

                if not request.POST.get('fi', '') or not request.POST.get('ff', ''):
                    raise NameError(f'Seleccione la fecha inicio/fin de la actividad.')
                fi, ff = convertir_fecha_hora_invertida(request.POST['fi'] + " 00:01"), convertir_fecha_hora_invertida(request.POST['ff'] + " 23:59")
                if not contrato:
                    raise NameError(f"Estimad{'o' if persona.es_mujer() else 'a'} {persona}, no se encontró un contrato vigente.")

                sec_informe = contrato.secuencia_informe()
                if SolicitudPago.objects.values('id').filter(status=True,fechainicio__date=fi.date(), fechaifin__date=ff.date(),
                                                             contrato=contrato).exists():
                    raise NameError(f"Estimad{'o' if persona.es_mujer() else 'a'} {persona}, ya cuenta con una solicitud en las fechas indicadas, favor de revisar en informes generados de posgrado.")
                if SolicitudPago.objects.values('id').filter(status=True, numero=sec_informe - 1,
                                                             fechainicio__date=fi.date(), fechaifin__date=ff.date(),
                                                             contrato=contrato).exists():
                    solicitud = SolicitudPago.objects.filter(status=True, numero=sec_informe - 1,
                                                             fechainicio__date=fi.date(), fechaifin__date=ff.date(),
                                                             contrato=contrato).order_by('-id').first()
                else:
                    solicitud = SolicitudPago(
                        fechainicio=fi,
                        fechaifin=ff,
                        contrato=contrato,
                        estado=0,
                        numero=sec_informe
                    )
                    solicitud.save(request)
                nombresinciales = ''
                nombre = persona.nombres.split()
                if len(nombre) > 1:
                    nombresiniciales = '{}{}'.format(nombre[0][0], nombre[1][0])
                else:
                    nombresiniciales = '{}'.format(nombre[0][0])
                inicialespersona = '{}{}{}'.format(nombresiniciales, persona.apellido1[0], persona.apellido2[0])
                name_file = f'informe_actividad_diaria_{inicialespersona.lower()}_{solicitud.numero}.pdf'
                data['fi'], data['ff'] = fi, ff
                fecha_actual = fi
                data['fechas'] = [fecha_actual + timedelta(days=d) for d in range((ff - fi).days + 1)]
                data['responsable'] = u"%s" % contrato.gestion.responsable
                data['contrato'] = contrato
                data['actividades'] = actividadesbitacora
                data['carreras'] = ContratoCarrera.objects.filter(contrato=contrato, status=True)
                data['areasprogramas'] = ContratoAreaPrograma.objects.filter(contrato=contrato, status=True)
                resp = conviert_html_to_pdf_name_bitacora(
                    'th_hojavida/informe_actividad_mensual_administrativo_posgrado.html',
                    {"data": data}, name_file)
                if resp[0]:
                    resp[1].seek(0)
                    fil_content = resp[1].read()
                    resp = ContentFile(fil_content)
                else:
                    return resp[1]
                requisito, hist_ =  requisito = solicitud.guardar_informe_solicitud_pago(request,requisito_id = 14,observacion = f'Informe generado por {persona.__str__()}',hoy = hoy,persona = persona,observacion_historial = f'Informe firmado por {persona.__str__()}',name_file = name_file, resp = resp)
                archivo_borrar = hist_.archivo.url
                texto = f'{persona}'
                random_number = random.randint(1, 1000000)
                data['archivo'] = hist_.archivo.url
                archivo = f"{hist_.archivo.url}?cache={random_number}"
                data['url_archivo'] = '{}{}'.format(dominio_sistema, archivo)
                data['id_objeto'] = hist_.id
                data['action_firma'] = 'firmainformemensualposgrado'
                cuerpo = f"Informe mensual de posgrado generado por {persona}"

                obshisto = HistorialObseracionSolicitudPago(
                    solicitud=solicitud,
                    observacion=cuerpo,
                    persona=persona,
                    estado=0,
                    fecha=hoy
                )
                obshisto.save(request)
                log(f'Genero el informe mensual de la solicitud: {solicitud}', request, 'change')
                template = get_template("formfirmaelectronica.html")
                return JsonResponse({"result": True, 'data': template.render(data), 'tienetoken': False})

                return JsonResponse({'result': False, 'to': '/pro_solicitudpago'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': False, 'mensaje': f'Error de conexión. {ex.__str__()}'})

        # DECIMO
        elif action == 'adddecimo':
            try:
                with transaction.atomic():
                    form = RegistraDecimoForm(request.POST, instancia=persona)
                    if not form.is_valid():
                        transaction.set_rollback(True)
                        form_error = [{k: v[0]} for k, v in form.errors.items()]
                        return JsonResponse({'result': True, "form": form_error, "mensaje": "Error en el formulario"})

                    ultimoregistro=RegistroDecimo.objects.filter(persona=persona,status=True).order_by('id').last()
                    if ultimoregistro:
                        if ultimoregistro.seleccion == int(form.cleaned_data['tiporegistro']):
                            return JsonResponse({'result': True,
                                                 "mensaje": f"No puede actualizar ya que actualmente {ultimoregistro.get_seleccion_display()} su décimo"})
                        ultimoregistro.fechafin=datetime.now().date()
                        ultimoregistro.activo=False
                        ultimoregistro.save(update_fields=['fechafin','activo'])

                    registro = RegistroDecimo(
                                        persona=persona,
                                        estado=1,
                                        seleccion=form.cleaned_data['tiporegistro'],
                                        fechainicio=datetime.now().date(),
                                        activo=True
                                        )
                    registro.save(request)

                    log(f'Agrego registro de decimo: {registro.get_seleccion_display()}', request, 'add')
                    return JsonResponse({'result': False, 'mensaje': 'Guardado con éxito'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": f"{ex}"})

        elif action == 'firmarcartadecimo':
            try:
                certificado = request.FILES["firma"]
                contrasenaCertificado = request.POST['palabraclave']
                razon = request.POST['razon'] if 'razon' in request.POST else ''
                extension_certificado = os.path.splitext(certificado.name)[1][1:]
                bytes_certificado = certificado.read()
                decimo = RegistroDecimo.objects.get(id=encrypt_id(request.POST['id']))
                firmas = []
                archivo_, url_archivo= generar_carta_decimos(request, decimo)
                palabras = f'{persona.nombre_completo_minus()} {persona.identificacion()}'.strip()
                x, y, numPage = obtener_posicion_x_y_saltolinea(url_archivo, palabras, True, True)
                if x and y:
                    firmas.append({'x': x, 'y': y, 'numPage': numPage})
                for membrete in firmas:
                    datau = JavaFirmaEc(
                        archivo_a_firmar=archivo_, archivo_certificado=bytes_certificado, extension_certificado=extension_certificado,
                        password_certificado=contrasenaCertificado,
                        page=int(membrete["numPage"]), reason=razon, lx=membrete["x"], ly=membrete["y"]
                    ).sign_and_get_content_bytes()
                    archivo_ = io.BytesIO()
                    archivo_.write(datau)
                    archivo_.seek(0)

                _name = f"carta_decimo_{persona.usuario}_{decimo.id}"
                file_obj = DjangoFile(archivo_, name=f"{_name}.pdf")

                decimo.archivo = file_obj
                decimo.estado = 2
                decimo.save(request)
                log(u'Firmo Documento: {}'.format(archivo_), request, "add")
                return JsonResponse({'result': False, 'mensaje': 'Guardado correctamente'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, "mensaje": 'Error: {}'.format(ex)}, safe=False)

        elif action == 'subircartadecimo':
            try:
                archivo = request.FILES["archivo"]
                if not archivo:
                    raise NameError('Por favor cargue un archivo antes de guardar.')
                decimo = RegistroDecimo.objects.get(id=encrypt_id(request.POST['id_obj']))
                decimo.archivo = archivo
                decimo.estado = 2
                decimo.save(request)
                log(u'Subo archivo firmado:  {}'.format(decimo), request, "edit")
                return JsonResponse({'result': False, 'mensaje': 'Guardado correctamente'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, "mensaje": 'Error: {}'.format(ex)}, safe=False)

        # PAZ Y SALVO
        elif action == 'addpazsalvo':
            try:
                with transaction.atomic():
                    form = PazSalvoFormHV(request.POST, instancia=persona)
                    formato = FormatoPazSalvo.objects.filter(status=True, activo=True).first()
                    if not form.is_valid():
                        transaction.set_rollback(True)
                        form_error = [{k: v[0]} for k, v in form.errors.items()]
                        return JsonResponse({'result': True, "form": form_error, "mensaje": "Error en el formulario"})

                    if not formato:
                        raise NameError('No existe un formato activado de paz y salvo para generar la solicitud')

                    pazsalvo = PazSalvo(formato=formato,
                                        persona=persona,
                                        departamento=form.cleaned_data['departamento'],
                                        cargo=form.cleaned_data['cargo'],
                                        tiporelacion=form.cleaned_data['tiporelacion'],
                                        jefeinmediato=form.cleaned_data['jefeinmediato'],
                                        ultimaremuneracion=form.cleaned_data['ultimaremuneracion'],
                                        fecha=form.cleaned_data['fecha'],
                                        motivosalida=form.cleaned_data['motivosalida'],
                                        estado=1)
                    pazsalvo.save(request)

                    titulo = 'Certificado de paz y salvo pendiente de responder'
                    sexo = 'o' if pazsalvo.jefeinmediato.sexo.id == 2 else 'a'
                    observacion = f'Estimado{sexo} {pazsalvo.jefeinmediato.nombre_completo_minus()} existe un nuevo certificado de paz y salvo de {pazsalvo.persona.nombre_completo_minus()} que necesita ser llenada y firmada por su persona.'
                    noti(titulo, observacion, pazsalvo.jefeinmediato, None, f'/th_pazsalvo?s={pazsalvo.persona.cedula}', pazsalvo.pk, 2, 'sagest', PazSalvo, request)
                    for d in formato.responsables_dp():
                        if not d.persona.id == pazsalvo.jefeinmediato.id:
                            sexo = 'o' if d.persona.sexo.id == 2 else 'a'
                            observacion = f'Estimado{sexo} {d.persona.nombre_completo_minus()} existe un nuevo certificado de paz y salvo de {pazsalvo.persona.nombre_completo_minus()} que necesita ser llenada y firmada por su persona.'
                            noti(titulo, observacion, d.persona, None, f'/th_pazsalvo?s={pazsalvo.persona.cedula}', pazsalvo.pk, 2, 'sga-sagest', PazSalvo, request)

                    log(f'Agrego paz y salvo: {pazsalvo}', request, 'add')
                    return JsonResponse({'result': False, 'mensaje': 'Guardado con éxito'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": f"{ex}"})

        elif action == 'editpazsalvo':
            try:
                pazsalvo = PazSalvo.objects.get(id=encrypt_id(request.POST['id']))
                form = PazSalvoFormHV(request.POST, instancia=pazsalvo)
                if not form.is_valid():
                    transaction.set_rollback(True)
                    form_error = [{k: v[0]} for k, v in form.errors.items()]
                    return JsonResponse({'result': True, "form": form_error, "mensaje": "Error en el formulario"})
                pazsalvo.departamento = form.cleaned_data['departamento']
                pazsalvo.cargo = form.cleaned_data['cargo']
                pazsalvo.tiporelacion = form.cleaned_data['tiporelacion']
                pazsalvo.jefeinmediato = form.cleaned_data['jefeinmediato']
                pazsalvo.ultimaremuneracion = form.cleaned_data['ultimaremuneracion']
                pazsalvo.fecha = form.cleaned_data['fecha']
                pazsalvo.motivosalida=form.cleaned_data['motivosalida']
                pazsalvo.save(request)
                log(f'Edito paz y salvo: {pazsalvo}', request, 'edit')
                return JsonResponse({'result': False, 'mensaje': 'Guardado con éxito'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': str(ex)})

        elif action == 'deldecimo':
            with transaction.atomic():
                try:
                    instancia = RegistroDecimo.objects.get(pk=int(encrypt(request.POST['id'])))
                    instancia.status = False
                    instancia.activo = False
                    instancia.save(request)
                    ultimoregistro = RegistroDecimo.objects.filter(persona=persona, status=True).order_by('id').last()
                    if ultimoregistro:
                        if not ultimoregistro.activo:
                            ultimoregistro.activo = True
                            ultimoregistro.save(update_fields=['activo'])
                    log(u'Elimino paz salvo: %s' % instancia, request, "del")
                    res_json = {"error": False, "mensaje": 'Registro eliminado'}
                except Exception as ex:
                    res_json = {'error': True, "message": "Error: {}".format(ex)}
                return JsonResponse(res_json, safe=False)

        elif action == 'delpazsalvo':
            with transaction.atomic():
                try:
                    instancia = PazSalvo.objects.get(pk=int(encrypt(request.POST['id'])))
                    instancia.status = False
                    instancia.save(request)
                    log(u'Elimino paz salvo: %s' % instancia, request, "del")
                    res_json = {"error": False, "mensaje": 'Registro eliminado'}
                except Exception as ex:
                    res_json = {'error': True, "message": "Error: {}".format(ex)}
                return JsonResponse(res_json, safe=False)

        elif action == 'firmarpazsalvo':
            try:
                certificado = request.FILES["firma"]
                contrasenaCertificado = request.POST['palabraclave']
                razon = request.POST['razon'] if 'razon' in request.POST else ''
                extension_certificado = os.path.splitext(certificado.name)[1][1:]
                bytes_certificado = certificado.read()
                pazsalvo = PazSalvo.objects.get(id=encrypt_id(request.POST['id']))
                firmas = []
                certificado_firma = pazsalvo.documento()
                archivo_ = certificado_firma.archivo
                titulo = persona.titulacion_principal_senescyt_registro().titulo.abreviatura if persona.titulacion_principal_senescyt_registro() else ''
                # Nota: El punto al final es esencial para el funcionamiento de la firma masiva, es el diferenciador ubicado en el html del certificado.
                palabras = f'{titulo} {persona}.'.strip()
                x, y, numPage = obtener_posicion_x_y_saltolinea(archivo_.url, palabras)
                if x and y:
                    x = x + len(palabras) + 8
                    y = y + 5
                    firmas.append({'x': x, 'y': y, 'numPage': numPage})
                for membrete in firmas:
                    datau = JavaFirmaEc(
                        archivo_a_firmar=archivo_, archivo_certificado=bytes_certificado, extension_certificado=extension_certificado,
                        password_certificado=contrasenaCertificado,
                        page=int(membrete["numPage"]), reason=razon, lx=membrete["x"], ly=membrete["y"]
                    ).sign_and_get_content_bytes()
                    archivo_ = io.BytesIO()
                    archivo_.write(datau)
                    archivo_.seek(0)

                _name = f"certificado_{certificado_firma.pazsalvo.persona.usuario}"
                file_obj = DjangoFile(archivo_, name=f"{_name}.pdf")

                certificado_firma.archivo = file_obj
                certificado_firma.save(request)
                log(u'Firmo Documento: {}'.format(archivo_), request, "add")

                historial = HistorialCertificadoFirmaPS(certificadosalida=certificado_firma,
                                                        archivo=file_obj,
                                                        persona=persona,
                                                        cargo=persona.mi_cargo_administrativo(),
                                                        cantidadfirmas=len(firmas),
                                                        estado=2)
                historial.save(request)

                pazsalvo = certificado_firma.pazsalvo
                pazsalvo.estado = 4
                pazsalvo.save(request)
                log(u'Edito estado a finalizado de paz y salvo:  {}'.format(pazsalvo), request, "edit")
                return JsonResponse({'result': False, 'mensaje': 'Guardado correctamente'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, "mensaje": 'Error: {}'.format(ex)}, safe=False)

        elif action == 'firmarpazsalvo_new':
            try:
                certificado = request.FILES["firma"]
                contrasenaCertificado = request.POST['palabraclave']
                razon = request.POST['razon'] if 'razon' in request.POST else ''
                extension_certificado = os.path.splitext(certificado.name)[1][1:]
                bytes_certificado = certificado.read()
                pazsalvo = PazSalvo.objects.get(id=encrypt_id(request.POST['id']))
                certificado_firma = pazsalvo.documento()
                archivo_ = certificado_firma.ultimo_archivo_con_exclusion(persona).archivo
                jsonFirmas = json.loads(request.POST['txtFirmas'])
                for membrete in jsonFirmas:
                    datau = JavaFirmaEc(
                        archivo_a_firmar=archivo_, archivo_certificado=bytes_certificado, extension_certificado=extension_certificado,
                        password_certificado=contrasenaCertificado,
                        page=int(membrete["numPage"]), reason=razon, lx=membrete["x"], ly=membrete["y"]
                    ).sign_and_get_content_bytes()
                    archivo_ = io.BytesIO()
                    archivo_.write(datau)
                    archivo_.seek(0)

                _name = f"certificado_{certificado_firma.pazsalvo.persona.usuario}"
                file_obj = DjangoFile(archivo_, name=f"{_name}.pdf")

                certificado_firma.archivo = file_obj
                certificado_firma.save(request)
                log(u'Firmo Documento: {}'.format(archivo_), request, "add")

                historial = HistorialCertificadoFirmaPS(certificadosalida=certificado_firma,
                                                        archivo=file_obj,
                                                        persona=persona,
                                                        cargo=persona.mi_cargo_administrativo(),
                                                        cantidadfirmas=len(jsonFirmas),
                                                        estado=2)
                historial.save(request)

                pazsalvo = certificado_firma.pazsalvo
                pazsalvo.estado = 4
                pazsalvo.save(request)
                log(u'Edito estado a finalizado de paz y salvo:  {}'.format(pazsalvo), request, "edit")
                return JsonResponse({'result': False, 'mensaje': 'Guardado correctamente'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, "mensaje": 'Error: {}'.format(ex)}, safe=False)

        elif action == 'subirpazsalvo':
            try:
                archivo = request.FILES["archivo"]
                if not archivo:
                    raise NameError('Por favor cargue un archivo antes de guardar.')
                pazsalvo = PazSalvo.objects.get(id=encrypt_id(request.POST['id_obj']))
                certificado_firma = pazsalvo.documento()
                archivo._name = generar_nombre("certificado", unidecode(archivo._name))
                certificado_firma.archivo = archivo
                certificado_firma.save(request)

                archivo._name = generar_nombre("historial", unidecode(archivo._name))
                historial = HistorialCertificadoFirmaPS(certificadosalida=certificado_firma,
                                                        archivo=archivo,
                                                        persona=persona,
                                                        cargo=persona.mi_cargo_administrativo(),
                                                        cantidadfirmas=1,
                                                        estado=2)
                historial.save(request)

                pazsalvo = certificado_firma.pazsalvo
                pazsalvo.estado = 4
                pazsalvo.save(request)
                log(u'Edito estado a finalizado de paz y salvo:  {}'.format(pazsalvo), request, "edit")
                return JsonResponse({'result': False, 'mensaje': 'Guardado correctamente'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, "mensaje": 'Error: {}'.format(ex)}, safe=False)

        elif action == 'subirrequisitos':
            with transaction.atomic():
                try:
                    pazsalvo = PazSalvo.objects.get(pk=encrypt_id(request.POST['id']))
                    beneficiario = pazsalvo.persona.nombre_normal_minus()
                    for doc in pazsalvo.documentos_subidos():
                        if f'documento_{doc.id}' in request.FILES:
                            newfile = request.FILES[f'documento_{doc.id}']
                            extension = newfile._name.split('.')
                            tam = len(extension)
                            exte = extension[tam - 1]
                            if newfile.size > 4194304:
                                raise NameError(u"Error, el tamaño del archivo es mayor a 4 Mb.")
                            if not exte.lower() in ['pdf']:
                                raise NameError(u"Error, solo archivos .pdf")
                            newfile._name = generar_nombre(f"DocumentoPS_{doc.id}_{pazsalvo.id}", newfile._name)
                            doc.archivo = newfile
                            if doc.estados == 2:
                                doc.estados = 3
                                doc.f_correccion = hoy
                                pazsalvo.estado_requisito=3
                                pazsalvo.save(request)
                            doc.save(request)
                    return JsonResponse({'result': False, 'mensaje': u'Guardado con exito'}, safe=False)
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": True, "mensaje": '{} {}'.format(str(ex), str(sys.exc_info()[-1].tb_lineno))}, safe=False)

        elif action == 'subirrequisitosperiodotthh':
            with transaction.atomic():
                try:
                    personaperiodotthh = PersonaPeriodotthh.objects.get(pk=encrypt_id(request.POST['id']))
                    for doc in personaperiodotthh.documentos_subidos():
                        if f'documento_{doc.id}' in request.FILES:
                            newfile = request.FILES[f'documento_{doc.id}']
                            extension = newfile._name.split('.')
                            tam = len(extension)
                            exte = extension[tam - 1]
                            if newfile.size > 2194304:
                                raise NameError(u"Error, el tamaño del archivo es mayor a 2 Mb.")
                            if not exte.lower() in ['pdf']:
                                raise NameError(u"Error, solo archivos .pdf")
                            newfile._name = generar_nombre(f"DocumentoREQUI_{doc.id}_{personaperiodotthh.id}", newfile._name)
                            doc.archivo = newfile
                            if doc.estados == 2:
                                doc.estados = 3
                                doc.f_correccion = hoy
                                personaperiodotthh.estado_requisito=3
                                personaperiodotthh.save(request)
                            doc.save(request)
                    return JsonResponse({'result': False, 'mensaje': u'Guardado con exito'}, safe=False)
                except Exception as ex:
                    lineaerror = 'Error on line {}'.format(sys.exc_info()[-1].tb_lineno)
                    return JsonResponse(
                        {"result": "bad", "mensaje": u"Intentelo más tarde. {} {}".format(str(ex), lineaerror)},safe=False)

        #SANCIONES:
        elif action == 'respuestadescargo':
            with transaction.atomic():
                try:
                    per_sancion = PersonaSancion.objects.get(pk=encrypt_id(request.POST['id']))
                    listado = json.loads(request.POST['lista_items1'])
                    for item in listado:
                        descripcion = item['descripcion']
                        if item['name_archivo'] in request.FILES:
                            archivo = request.FILES[item['name_archivo']]
                            exte = ext_archive(archivo._name)
                            if archivo.size > 5194304:
                                raise NameError(f"El tamaño del archivo: {descripcion} es mayor a 5 Mb.")
                            if not exte in ['.png', '.jpg', '.jpeg', '.pdf']:
                                raise NameError(u"Solo se permite archivos con extensión .pdf, .jpg, .jpeg, .png")
                            name_archive = unidecode(descripcion[:8])
                            archivo._name = generar_nombre(f"{per_sancion.id}_{name_archive}", archivo._name)
                            res_descargo = RespuestaDescargo(personasancion=per_sancion, descripcion=descripcion, archivo=archivo)
                            res_descargo.save(request)
                    log(f'Agrego {len(listado)} respuestas de descargo: {per_sancion}', request, "add")
                    return JsonResponse({'result': False, 'mensaje': u'Guardado con exito'}, safe=False)
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": True, "mensaje": '{} {}'.format(str(ex), str(sys.exc_info()[-1].tb_lineno))}, safe=False)

        elif action == 'editrespuestadescargo':
            with transaction.atomic():
                try:
                    respuesta = RespuestaDescargo.objects.get(pk=encrypt_id(request.POST['id']))
                    form = RespuestaDescargoForm(request.POST, request.FILES, instancia=respuesta)
                    if not form.is_valid():
                        form_error = [{k: v[0]} for k, v in form.errors.items()]
                        return JsonResponse({'result': True, "form": form_error, "mensaje": "Error en el formulario"})
                    respuesta.archivo=form.cleaned_data['archivo']
                    respuesta.descripcion=form.cleaned_data['descripcion']
                    respuesta.save(request)
                    log(u'Elimino respuesta de descargo: %s' % respuesta, request, "del")
                    return JsonResponse({'result': False, 'mensaje': u'Guardado con exito'}, safe=False)
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": True, "mensaje": '{} {}'.format(str(ex), str(sys.exc_info()[-1].tb_lineno))}, safe=False)

        elif action == 'delrespuesta':
            with transaction.atomic():
                try:
                    respuesta = RespuestaDescargo.objects.get(pk=encrypt_id(request.POST['id']))
                    respuesta.status=False
                    respuesta.save(request)
                    log(u'Elimino respuesta de descargo: %s' % respuesta, request, "del")
                    return JsonResponse({'error': False, 'mensaje': u'Guardado con exito'})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"error": True, "mensaje": '{} {}'.format(str(ex), str(sys.exc_info()[-1].tb_lineno))})

        elif action == 'firmardocumento':
            try:
                context = {'lx': 395, 'ly_menos': 25}
                firmar_documento_etapa(request, persona, context)
                return JsonResponse({"result": False, "mensaje": "Guardado con exito", }, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error: %s" % ex})

        elif action == 'justificarnoasistencia':
            try:
                form = JustificacionPersonaAudienciaForm(request.POST, request.FILES)
                if not form.is_valid():
                    form_error = [{k: v[0]} for k, v in form.errors.items()]
                    return JsonResponse({'result': True, "form": form_error, "mensaje": "Error en el formulario"})
                persona_audiencia = PersonaAudienciaSancion.objects.get(pk=encrypt_id(request.POST['id']))
                persona_audiencia.justificacion = form.cleaned_data['justificacion']
                persona_audiencia.archivo = form.cleaned_data['archivo']
                persona_audiencia.asistira = False
                persona_audiencia.save(request)
                persona_sancion = persona_audiencia.get_persona_sancion()
                persona_sancion.bloqueo = False
                persona_sancion.save(request)
                # Necesario actualizar la sesión para liberar el sistema
                request.session['persona_sancion'] = PersonaSancion.objects.filter(persona=persona, status=True, bloqueo=True).first()
                log(u'Justifico inasistencia: %s' % persona_audiencia, request, "edit")
                return JsonResponse({'result': False, 'mensaje': 'Guardado con exito'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": 'Error: {}'.format(ex)})

        elif action == 'confirmarasistencia':
            try:
                persona_audiencia = PersonaAudienciaSancion.objects.get(pk=encrypt_id(request.POST['id']))
                persona_audiencia.asistira = True
                persona_audiencia.validacion_asis = 1
                persona_audiencia.justificacion = ''
                persona_audiencia.archivo = None
                persona_audiencia.save(request)
                persona_sancion = persona_audiencia.get_persona_sancion()
                persona_sancion.bloqueo = False
                persona_sancion.save(request)
                if persona_audiencia.audiencia.todos_asistiran():
                    audiencia = persona_audiencia.audiencia
                    audiencia.estado = 2
                    audiencia.save(request)
                # Necesario actualizar la sesión para liberar el sistema
                request.session['persona_sancion'] = PersonaSancion.objects.filter(persona=persona, status=True, bloqueo=True).first()
                log(u'Confirmo asistencia: %s' % persona_audiencia, request, "edit")
                return JsonResponse({'result': 'ok', 'mensaje': 'Guardado con exito'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": 'bad', "mensaje": 'Error: {}'.format(ex)})

        elif action == 'confirmardescargo':
            try:
                persna_sancion = PersonaSancion.objects.get(pk=encrypt_id(request.POST['id']))
                persna_sancion.bloqueo = False
                persna_sancion.save(request)
                # Necesario actualizar la sesión para liberar el sistema
                request.session['persona_sancion'] = PersonaSancion.objects.filter(persona=persona, status=True, bloqueo=True).first()
                log(u'Confirmo carga de respuestas de recargo: %s' % persna_sancion, request, "edit")
                return JsonResponse({'result': 'ok', 'mensaje': 'Guardado con exito'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": 'bad', "mensaje": 'Error: {}'.format(ex)})

        elif action == 'sifirmaraccionpers':
            try:
                obj = ConsultaFirmaPersonaSancion.objects.get(pk=encrypt_id(request.POST['id']))
                obj.estado = 1
                obj.fecha_respuesta = datetime.now()
                obj.save(request)
                persona_sancion = obj.persona_sancion
                persona_sancion.bloqueo = False
                persona_sancion.save(request)
                # Necesario actualizar la sesión para liberar el sistema
                request.session['persona_sancion'] = PersonaSancion.objects.filter(persona=persona, status=True, bloqueo=True).first()
                log(u'Confirmo firmar acción de personal: %s' % obj, request, "edit")
                return JsonResponse({'result': 'ok', 'mensaje': 'Guardado con exito'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": 'bad', "mensaje": 'Error: {}'.format(ex)})

        elif action == 'nofirmaraccionpers':
            try:
                form = MotivoNoFirmaAccionPersonalForm(request.POST)
                if not form.is_valid():
                    form_error = [{k: v[0]} for k, v in form.errors.items()]
                    return JsonResponse({'result': True, "form": form_error, "mensaje": "Error en el formulario"})
                obj = ConsultaFirmaPersonaSancion.objects.get(pk=encrypt_id(request.POST['id']))
                obj.estado = 2
                obj.fecha_respuesta = datetime.now()
                obj.motivo = form.cleaned_data['motivo']
                obj.save(request)
                persona_sancion = obj.persona_sancion
                persona_sancion.bloqueo = False
                persona_sancion.save(request)
                # Necesario actualizar la sesión para liberar el sistema
                request.session['persona_sancion'] = PersonaSancion.objects.filter(persona=persona, status=True, bloqueo=True).first()
                log(u'Rechaza firma de acción de personal: %s' % obj, request, "edit")
                return JsonResponse({'result': False})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": 'Error: {}'.format(ex)})


        return JsonResponse({"result": "bad", "mensaje": u"Solicitud Incorrecta."})
    else:
        if 'action' in request.GET:
            data['action'] = action = request.GET['action']


            if action == 'informe-administrativo-posgrado':
                import unicodedata
                try:
                    data['title'] = u"INFORME DE ACTIVIDADES ADMINISTRATIVAS"
                    data['hoy'] = hoy = datetime.now()
                    if not request.GET.get('fi', '') or not request.GET.get('ff', ''):
                        return HttpResponseRedirect(f"/th_hojavida?info=Seleccione la fecha inicio/fin de la actividad.")

                    fi, ff = convertir_fecha_hora_invertida(request.GET['fi'] + " 00:01"), convertir_fecha_hora_invertida(request.GET['ff'] + " 23:59")
                    actividadesbitacora = BitacoraActividadDiaria.objects.filter(Q(fecha__gte=fi, fecha__lte=ff, persona=persona, status=True)).order_by('fecha')
                    contrato = ContratoDip.objects.filter(Q(persona=persona, fechainicio__lte=fi, fechainicio__lt=ff, fechafin__gt=fi, fechafin__gte=ff,status=True)).order_by('-fecha_creacion').first()
                    if contrato:
                        data['fi'], data['ff'] = fi, ff
                        fecha_actual = fi
                        data['fechas'] = [fecha_actual + timedelta(days=d) for d in range((ff - fi).days + 1)]
                        data['responsable'] = u"%s" % contrato.seccion.responsable
                        data['contrato'] = contrato
                        data['actividades'] = actividadesbitacora
                        data['carreras'] = ContratoCarrera.objects.filter(contrato=contrato, status=True)
                        data['areasprogramas'] = ContratoAreaPrograma.objects.filter(contrato=contrato, status=True)
                        nombres = unicodedata.normalize('NFD', u"%s" % persona.un_nombre_dos_apellidos()).encode('ascii', 'ignore').decode("utf-8").upper()
                        return conviert_html_to_pdf_name('th_hojavida/informe_actividad_mensual_administrativo_posgrado.html', {"data": data},f"INFORME ADMINISTRATIVO - {nombres} {hoy.hour}{hoy.minute}{hoy.second}")
                    else:
                        return HttpResponseRedirect(f"/th_hojavida?info=Estimad{'o' if persona.es_mujer() else 'a'} {persona}, no se encontró un contrato vigente.")
                except Exception as ex:
                    return HttpResponseRedirect(f"/th_hojavida?info=Error de conexión. {ex.__str__()}")

            if action == 'bloquear':
                try:
                    if 'id' in request.GET:
                        nivel = NivelTitulacion.objects.get(pk=int(request.GET['id']))
                        if nivel.rango == 6:
                            data = {"results": "ok", "rango": 1}
                        else:
                            data = {"results": "ok", "rango": 2}
                        return JsonResponse(data)
                except Exception as ex:
                    pass

            elif action == 'subirfoto':
                try:
                    form = PersonaSubirFotoForm()
                    data['title'] = u'Actualizar Foto'
                    data['form'] = form
                    template = get_template("th_hojavida/subirfoto.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content, 'title': data['title']})
                except Exception as ex:
                    pass

            elif action == 'editfoto':
                try:
                    template = get_template("th_hojavida/modal/editfoto.html")
                    return JsonResponse({'result': True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({'result': False, 'mensaje': f'Error: {ex}'})

            elif action == 'editrevista':
                try:
                    data['title'] = u'Editar Revista' if int(request.GET['tiporegistro']) == 1 else 'Editar Congreso'
                    data['revista'] = revista = RevistaInvestigacion.objects.get(pk=request.GET['id'])
                    basesindexadas = BaseIndexadaInvestigacion.objects.filter(revistainvestigacionbase__revista=revista,
                                                                              revistainvestigacionbase__status=True).order_by(
                        'nombre')

                    form = RevistaInvestigacionHojaVidaForm(initial={'nombrerevista2': revista.nombre,
                                                                     'codigoissn2': revista.codigoissn,
                                                                     'institucion2': revista.institucion,
                                                                     'tipo2': revista.tipo,
                                                                     'enlacerevista2': revista.enlace,
                                                                     'baseindexada2': basesindexadas
                                                                     })
                    data['form'] = form
                    template = get_template("th_hojavida/editrevista.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content, 'title': data['title']})
                except Exception as ex:
                    pass

            elif action == 'addpublicacion':
                try:
                    data['title'] = u'Adicionar Solicitud Publicación'

                    form = SolicitudPublicacionForm()
                    form2 = RevistaInvestigacionForm()
                    data['form'] = form
                    data['form2'] = form2
                    return render(request, "th_hojavida/addarticulo.html", data)
                except Exception as ex:
                    pass

            elif action == 'addcapitulolibro':
                try:
                    data['title'] = u'Adicionar Capítulo de Libro'
                    form = CapituloLibroHojaVidaForm()
                    data['form'] = form
                    return render(request, "th_hojavida/addcapitulolibro.html", data)
                except Exception as ex:
                    pass


            elif action == 'delparrevisor':
                try:
                    data['title'] = u'Elimina Revisión Artículo'
                    data['parrevisor'] = ParRevisorProduccionCientifica.objects.get(pk=request.GET['id'])
                    return render(request, "th_hojavida/delparrevision.html", data)
                except Exception as ex:
                    pass

            elif action == 'deleteAportacionHistorialLaboral':
                try:
                    data['title'] = u'Elimina Historial Laboral'
                    data['registro'] = PersonaAportacionHistorialLaboral.objects.get(pk=int(encrypt(request.GET['id'])))
                    return render(request, "th_hojavida/delaportacionhistoriallaboral.html", data)
                except Exception as ex:
                    pass

            # elif action == 'editparrevisor':
            #     try:
            #         data['title'] = u'Editar Revisión de Artículo'
            #         data['parrevision'] = parrevision = ParRevisorProduccionCientifica.objects.get(pk=request.GET['id'])
            #         initial = model_to_dict(parrevision)
            #         form2 = RevistaInvestigacionForm()
            #         data['form'] = ParRevisorProduccionCientificaForm(initial=initial)
            #         data['form2'] = form2
            #         return render(request, "th_hojavida/editparrevision.html", data)
            #     except Exception as ex:
            #         pass

            elif action == 'malladistributivo':
                try:
                    data['admisionperiodo'] = periodo = request.session['periodoadmision']
                    carrera = Carrera.objects.get(pk=int(request.GET['id']))
                    mallas = Malla.objects.filter(
                        pk__in=Materia.objects.values_list('asignaturamalla__malla').filter(nivel__periodo=periodo,
                                                                                            asignaturamalla__malla__carrera=carrera).distinct()).distinct()
                    lista = []
                    for ca in mallas:
                        lista.append([ca.id, "%s" % ca])
                    return JsonResponse({'result': 'ok', 'lista': lista})
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

            elif action == 'carreradistributivo':
                try:
                    data['admisionperiodo'] = periodo = request.session['periodoadmision']
                    coordinacion = Coordinacion.objects.get(pk=int(request.GET['id']))
                    carreras = Carrera.objects.filter(
                        pk__in=Materia.objects.values_list('asignaturamalla__malla__carrera').exclude(
                            asignaturamalla__malla__carrera__modalidad=3).filter(nivel__periodo=periodo,
                                                                                 asignaturamalla__malla__carrera__coordinacion=coordinacion).distinct()).distinct()
                    lista = []
                    for ca in carreras:
                        lista.append([ca.id, "%s" % ca])
                    return JsonResponse({'result': 'ok', 'lista': lista})
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

            elif action == 'datospersonalesaspirante':
                try:
                    data['title'] = u'Datos personales'
                    persona = Persona.objects.get(id=persona.id)
                    maternidad = persona.personadetallematernidad_set.filter(
                        Q(status_gestacion=True) | Q(status_lactancia=True)).first()
                    if persona.sexo:
                        if persona.sexo.id == 1 and maternidad:
                            data['form'] = PersonaDetalleMaternidadAspiranteForm(
                                initial={'estadogestacion': maternidad.status_gestacion,
                                         'semanasembarazo': maternidad.semanasembarazo,
                                         'lactancia': maternidad.status_lactancia,
                                         'fechaparto': maternidad.fechaparto})
                        elif not maternidad:
                            data['form'] = PersonaDetalleMaternidadAspiranteForm()

                    form = DatosPersonalesAspiranteForm(initial={'nombres': persona.nombres,
                                                                 'apellido1': persona.apellido1,
                                                                 'apellido2': persona.apellido2,
                                                                 'cedula': persona.cedula,
                                                                 'pasaporte': persona.pasaporte,
                                                                 'extension': persona.telefonoextension,
                                                                 'sexo': persona.sexo,
                                                                 'lgtbi': persona.lgtbi,
                                                                 'anioresidencia': persona.anioresidencia,
                                                                 'nacimiento': persona.nacimiento,
                                                                 'email': persona.email,
                                                                 'estadocivil': persona.estado_civil(),
                                                                 'libretamilitar': persona.libretamilitar,
                                                                 'eszurdo': persona.eszurdo,
                                                                 'estadogestacion': persona.estadogestacion,
                                                                 })
                    form.editar()
                    if esestudiante:
                        form.es_estudiante()
                    data['form'] = form
                    data['persona'] = persona
                    banderalibreta = 0
                    banderapapeleta = 0
                    banderacedula = 0
                    documentos = PersonaDocumentoPersonal.objects.filter(persona=persona)
                    if documentos:
                        if documentos[0].libretamilitar:
                            banderalibreta = 1
                        if documentos[0].papeleta:
                            banderapapeleta = 1
                        if documentos[0].cedula:
                            banderacedula = 1

                    data['banderacedula'] = banderacedula
                    data['banderalibreta'] = banderalibreta
                    data['banderapapeleta'] = banderapapeleta
                    return render(request, "th_hojavida/datospersonalesaspirante.html", data)
                except Exception as ex:
                    pass


            elif action == 'datosinstitucionales':
                try:
                    data['title'] = u'Datos Institucionales'
                    form = DatosInstitucionalesPersonaForm(initial={'extension': persona.telefonoextension})
                    data['form'] = form
                    return render(request, "th_hojavida/datosinstitucionales.html", data)
                except Exception as ex:
                    pass

            elif action == 'datosnacimientoaspirante':
                try:
                    data['title'] = u'Datos de nacimiento'
                    data['nacionalidad'] = persona.paisnacimiento.nacionalidad
                    form = DatosNacimientoAspiranteForm(initial={'paisnacimiento': persona.paisnacimiento,
                                                                 'provincianacimiento': persona.provincianacimiento,
                                                                 'cantonnacimiento': persona.cantonnacimiento,
                                                                 'parroquianacimiento': persona.parroquianacimiento,
                                                                 'nacionalidad': persona.nacionalidad})
                    form.editar(persona)
                    data['form'] = form
                    return render(request, "th_hojavida/datosnacimientoaspirante.html", data)
                except Exception as ex:
                    pass

            elif action == 'datossituacionlaboralaspirante':
                try:
                    data['title'] = u'Datos de Situacion Laboral'
                    data['action'] = action
                    form = DatosSituacionLaboralAspiranteForm()
                    if persona.personasituacionlaboral_set.filter(status=True).exists():
                        data['situacion'] = situacion = PersonaSituacionLaboral.objects.get(persona=persona)
                        form = DatosSituacionLaboralAspiranteForm(initial=model_to_dict(situacion))
                    data['form'] = form
                    return render(request, "th_hojavida/datossituacionlaboraladmision.html", data)
                except Exception as ex:
                    pass

            # elif action == 'datosmedicos':
            #     try:
            #         data['title'] = u'Datos medicos basicos'
            #         datosextension = persona.datos_extension()
            #         examenfisico = persona.datos_examen_fisico()
            #         form = DatosMedicosForm(initial={'carnetiess': datosextension.carnetiess,
            #                                          'sangre': persona.sangre,
            #                                          'peso': examenfisico.peso,
            #                                          'talla': examenfisico.talla})
            #         # if not perfilprincipal.es_estudiante():
            #         #     form.deshabilitar()
            #         data['form'] = form
            #         banderatiposangre = 0
            #         documentos = PersonaDocumentoPersonal.objects.filter(persona=persona)
            #         if documentos:
            #             if documentos[0].tiposangre:
            #                 banderatiposangre = 1
            #
            #         data['banderatiposangre'] = banderatiposangre
            #
            #         return render(request, "th_hojavida/datosmedicos.html", data)
            #     except Exception as ex:
            #         pass

            elif action == 'datosdomicilioaspirante':
                try:
                    data['title'] = u'Datos de domicilio'
                    form = DatosDomicilioAdmisionForm(initial={'pais': persona.pais,
                                                               'provincia': persona.provincia,
                                                               'canton': persona.canton,
                                                               'ciudadela': persona.ciudadela,
                                                               'parroquia': persona.parroquia,
                                                               'direccion': persona.direccion,
                                                               'direccion2': persona.direccion2,
                                                               'sector': persona.sector,
                                                               'num_direccion': persona.num_direccion,
                                                               'referencia': persona.referencia,
                                                               'telefono': persona.telefono,
                                                               'telefono_conv': persona.telefono_conv,
                                                               'tipocelular': persona.tipocelular,
                                                               'zona': persona.zona})
                    form.editar(persona)
                    data['form'] = form
                    return render(request, "th_hojavida/datosdomicilioaspirante.html", data)
                except Exception as ex:
                    pass

            elif action == 'contactoemergencia':
                try:
                    data['title'] = u'Contacto de emergencia'
                    datosextension = persona.datos_extension()
                    form = ContactoEmergenciaForm(initial={'contactoemergencia': datosextension.contactoemergencia,
                                                           'parentescoemergencia': datosextension.parentescoemergencia,
                                                           'telefonoemergencia': datosextension.telefonoemergencia})
                    data['form'] = form
                    return render(request, "th_hojavida/contactoemergencia.html", data)
                except Exception as ex:
                    pass

            elif action == 'contactoemergenciaaspirante':
                try:
                    data['title'] = u'Contacto de emergencia'
                    datosextension = persona.datos_extension()
                    form = ContactoEmergenciaAspiranteForm(
                        initial={'contactoemergencia': datosextension.contactoemergencia,
                                 'parentescoemergencia': datosextension.parentescoemergencia,
                                 'telefonoemergencia': datosextension.telefonoemergencia})
                    data['form'] = form
                    return render(request, "th_hojavida/contactoemergenciaaspirante.html", data)
                except Exception as ex:
                    pass

            elif action == 'etnia':
                try:
                    data['title'] = u'Etnia/Pueblo/Nacionalidad'
                    perfil = persona.mi_perfil()
                    form = EtniaForm(initial={'raza': perfil.raza,
                                              'nacionalidadindigena': perfil.nacionalidadindigena})
                    tienearchivo = True if perfil.archivoraza else False
                    data['form'] = form
                    data['tienearchivo'] = tienearchivo
                    return render(request, "th_hojavida/etnia.html", data)
                except Exception as ex:
                    pass

            elif action == 'etniaaspirante':
                try:
                    data['title'] = u'Etnia/Pueblo/Nacionalidad'
                    perfil = persona.mi_perfil()
                    form = EtniaAspiranteForm(initial={'raza': perfil.raza,
                                                       'nacionalidadindigena': perfil.nacionalidadindigena})
                    tienearchivo = True if perfil.archivoraza else False
                    data['form'] = form
                    data['tienearchivo'] = tienearchivo
                    return render(request, "th_hojavida/etniaaspirante.html", data)
                except Exception as ex:
                    pass

            elif action == 'discapacidadaspirante':
                from sga.models import Discapacidad
                try:
                    data['title'] = u'Discapacidad'
                    perfil = persona.mi_perfil()
                    form = DiscapacidadAspiranteForm(initial=model_to_dict(perfil))
                    tienearchivo = True if perfil.archivo else False
                    data['form'] = form
                    data['tienearchivo'] = tienearchivo
                    data['tipodis'] = Discapacidad.objects.filter(status=True)
                    return render(request, "th_hojavida/discapacidadaspirante.html", data)
                except Exception as ex:
                    pass

            elif action == 'addfamiliaraspirante':
                try:
                    data['title'] = u'Adicionar familiar'
                    data['form'] = FamiliarAspiranteForm()
                    return render(request, "th_hojavida/addfamiliaraspirante.html", data)
                except Exception as ex:
                    pass

            elif action == 'editfamiliaraspirante':
                try:
                    data['title'] = u'Editar familiar'
                    data['familiar'] = familiar = PersonaDatosFamiliares.objects.get(pk=int(request.GET['id']))
                    apellido1, apellido2, nombres = '', '', familiar.nombre
                    if familiar.personafamiliar:
                        apellido1 = familiar.personafamiliar.apellido1
                        apellido2 = familiar.personafamiliar.apellido2
                        nombres = familiar.personafamiliar.nombres
                    banderacedula = 0
                    if familiar.cedulaidentidad:
                        banderacedula = 1
                    data['banderacedula'] = banderacedula
                    form = FamiliarAspiranteForm(initial={'identificacion': familiar.identificacion,
                                                          'parentesco': familiar.parentesco,
                                                          'nombre': nombres,
                                                          'apellido1': apellido1,
                                                          'apellido2': apellido2,
                                                          'nacimiento': familiar.nacimiento,
                                                          'fallecido': familiar.fallecido,
                                                          'tienediscapacidad': familiar.tienediscapacidad,
                                                          'telefono': familiar.telefono,
                                                          'niveltitulacion': familiar.niveltitulacion,
                                                          'ingresomensual': familiar.ingresomensual,
                                                          'formatrabajo': familiar.formatrabajo,
                                                          'telefono_conv': familiar.telefono_conv,
                                                          'trabajo': familiar.trabajo,
                                                          'convive': familiar.convive,
                                                          'sustentohogar': familiar.sustentohogar,
                                                          'essustituto': familiar.essustituto,
                                                          'autorizadoministerio': familiar.autorizadoministerio,
                                                          'tipodiscapacidad': familiar.tipodiscapacidad,
                                                          'porcientodiscapacidad': familiar.porcientodiscapacidad,
                                                          'carnetdiscapacidad': familiar.carnetdiscapacidad,
                                                          'institucionvalida': familiar.institucionvalida,
                                                          'tipoinstitucionlaboral': familiar.tipoinstitucionlaboral,
                                                          'tienenegocio': familiar.tienenegocio,
                                                          'negocio': familiar.negocio, })
                    form.edit()
                    data['form'] = form
                    return render(request, "th_hojavida/editfamiliaraspirante.html", data)
                except Exception as ex:
                    pass

            elif action == 'delfamiliar':
                try:
                    data['title'] = u'Eliminar familiar'
                    data['familiar'] = familiar = PersonaDatosFamiliares.objects.get(pk=int(request.GET['id']))
                    return render(request, "th_hojavida/delfamiliar.html", data)
                except Exception as ex:
                    pass


            elif action == 'adddeclaracionaspirante':
                try:
                    data['title'] = u'Adicionar declaración'
                    form = DeclaracionBienAspiranteForm()
                    form.ocultarcampos()
                    data['form'] = form
                    return render(request, "th_hojavida/adddeclaracionaspirante.html", data)
                except Exception as ex:
                    pass

            elif action == 'deldeclaracion':
                try:
                    data['title'] = u'Eliminar declaración'
                    data['declaracion'] = declaracion = DeclaracionBienes.objects.get(pk=int(request.GET['id']))
                    return render(request, "th_hojavida/deldeclaracion.html", data)
                except Exception as ex:
                    pass




            elif action == 'delvacunacion':
                try:
                    data['title'] = u'Eliminar Evidencia de Vacunación'
                    data['vacuna'] = vacuna = VacunaCovid.objects.get(pk=int(request.GET['id']))
                    return render(request, "th_hojavida/delvacunacion.html", data)
                except Exception as ex:
                    pass

            elif action == 'addtitulacionaspirante':
                try:
                    data['title'] = u'Adicionar titulación'
                    form = TitulacionPersonaAdmisionForm()
                    form.adicionar()
                    data['form'] = form
                    return render(request, "th_hojavida/addtitulacionaspirante.html", data)
                except Exception as ex:
                    pass

            elif action == 'edittitulacionaspirante':
                try:
                    data['title'] = u'Editar titulación'
                    data['titulacion'] = titulacion = Titulacion.objects.get(pk=int(request.GET['id']))
                    campotitulo, campoamplio, campoespecifico, campodetallado = None, None, None, None
                    if CamposTitulosPostulacion.objects.filter(status=True, titulo=titulacion.titulo).exists():
                        campotitulo = CamposTitulosPostulacion.objects.filter(status=True,
                                                                              titulo=titulacion.titulo).first()
                        campoamplio = AreaConocimientoTitulacion.objects.filter(status=True,
                                                                                id__in=campotitulo.campoamplio.all().values_list(
                                                                                    'id', flat=True))
                        campoespecifico = SubAreaConocimientoTitulacion.objects.filter(status=True,
                                                                                       id__in=campotitulo.campoespecifico.all().values_list(
                                                                                           'id', flat=True))
                        campodetallado = SubAreaEspecificaConocimientoTitulacion.objects.filter(status=True,
                                                                                                id__in=campotitulo.campodetallado.all().values_list(
                                                                                                    'id', flat=True))
                    form = TitulacionPersonaAdmisionForm(initial={'titulo': titulacion.titulo,
                                                                  'areatitulo': titulacion.areatitulo,
                                                                  'fechainicio': titulacion.fechainicio,
                                                                  'educacionsuperior': titulacion.educacionsuperior,
                                                                  'institucion': titulacion.institucion,
                                                                  'colegio': titulacion.colegio,
                                                                  'cursando': titulacion.cursando,
                                                                  'fechaobtencion': titulacion.fechaobtencion if not titulacion.cursando else datetime.now().date(),
                                                                  'fechaegresado': titulacion.fechaegresado if not titulacion.cursando else datetime.now().date(),
                                                                  'registro': titulacion.registro,
                                                                  # 'areaconocimiento': titulacion.areaconocimiento,
                                                                  # 'subareaconocimiento': titulacion.subareaconocimiento,
                                                                  # 'subareaespecificaconocimiento': titulacion.subareaespecificaconocimiento,
                                                                  'pais': titulacion.pais,
                                                                  'provincia': titulacion.provincia,
                                                                  'canton': titulacion.canton,
                                                                  'parroquia': titulacion.parroquia,
                                                                  'campoamplio': campoamplio,
                                                                  'campoespecifico': campoespecifico,
                                                                  'campodetallado': campodetallado,
                                                                  'anios': titulacion.anios,
                                                                  'semestres': titulacion.semestres,
                                                                  'aplicobeca': titulacion.aplicobeca,
                                                                  'tipobeca': titulacion.tipobeca,
                                                                  'financiamientobeca': titulacion.financiamientobeca,
                                                                  'valorbeca': titulacion.valorbeca})
                    form.editar(titulacion)
                    data['form'] = form
                    return render(request, "th_hojavida/edittitulacionaspirante.html", data)
                except Exception as ex:
                    pass

            elif action == 'addarchivotitulacion':
                try:
                    data['title'] = u'Adicionar archivo de titulación'
                    data['titulacion'] = titulacion = Titulacion.objects.get(pk=int(request.GET['id']))
                    data['form'] = ArchivoTitulacionForm()
                    return render(request, "th_hojavida/addarchivotitulacion.html", data)
                except Exception as ex:
                    pass

            elif action == 'addarchivoidioma':
                try:
                    data['title'] = u'Adicionar archivo de idiomas'
                    data['idioma'] = idioma = IdiomaDomina.objects.get(pk=int(request.GET['id']))
                    data['form'] = ArchivoIdiomaForm()
                    return render(request, "th_hojavida/addarchivoidioma.html", data)
                except Exception as ex:
                    pass

            elif action == 'addarchivocapacitacion':
                try:
                    data['title'] = u'Adicionar archivo de capacitación'
                    data['capacitacion'] = capacitacion = Capacitacion.objects.get(pk=int(request.GET['id']))
                    data['form'] = ArchivoCapacitacionForm()
                    return render(request, "th_hojavida/addarchivocapacitacion.html", data)
                except Exception as ex:
                    pass

            elif action == 'addarchivoexperiencia':
                try:
                    data['title'] = u'Adicionar archivo de referencia laboral'
                    data['experiencia'] = experiencia = ExperienciaLaboral.objects.get(pk=int(request.GET['id']))
                    data['form'] = ArchivoExperienciaForm()
                    return render(request, "th_hojavida/addarchivoexperiencia.html", data)
                except Exception as ex:
                    pass

            elif action == 'addarchivocontratoper':
                try:
                    if HistorialArchivosContratos.objects.filter(personacontrato_id=int(request.GET['id']),
                                                                 status=True):
                        if not HistorialArchivosContratos.objects.filter(personacontrato_id=int(request.GET['id']),
                                                                         estado_archivo=4, status=True):
                            return JsonResponse({"result": "bad",
                                                 "mensaje": u"Ya subió archivo de contrato firmado. Revisar proceso en apartado info."})

                    data['title'] = u'Adicionar archivo firmado de contrato personal'
                    data['contrato'] = contrato = PersonaContratos.objects.get(pk=int(request.GET['id']))
                    data['form'] = ArchivoHistorialContratoForm()
                    return render(request, "th_hojavida/addarchivohistorialcontrato.html", data)
                except Exception as ex:
                    pass

            elif action == 'addidioma':
                try:
                    data['title'] = u'Adicionar idioma'
                    form = IdiomaDominaForm()
                    data['form'] = form
                    return render(request, "th_hojavida/addidioma.html", data)
                except Exception as ex:
                    pass

            elif action == 'editidioma':
                try:
                    data['title'] = u'Editar Idioma'
                    data['idioma'] = idioma = IdiomaDomina.objects.get(pk=request.GET['id'])
                    data['form'] = IdiomaDominaForm(initial={'idioma': idioma.idioma,
                                                             'escritura': idioma.escritura,
                                                             'oral': idioma.oral,
                                                             'lenguamaterna': idioma.lenguamaterna,
                                                             'lectura': idioma.lectura})
                    return render(request, "th_hojavida/editidioma.html", data)
                except Exception as ex:
                    pass

            elif action == 'delidioma':
                try:
                    data['title'] = u'Eliminar idioma'
                    data['idioma'] = idioma = IdiomaDomina.objects.get(pk=int(request.GET['id']))
                    return render(request, "th_hojavida/delidioma.html", data)
                except Exception as ex:
                    pass

            elif action == 'addgasto':
                try:
                    data['title'] = u'Adiciomnar Gastos Personales'
                    hoy = datetime.now().date()
                    fechaingreso = persona.mi_fechaingreso()
                    data['gastospersonales'] = GastosPersonales.objects.filter(persona=persona,
                                                                               periodogastospersonales__anio=datetime.now().date().year,
                                                                               status=True).order_by('anio')
                    data['rolpago_ingresos'] = persona.rolpago_ingresos()
                    data['rmu_actual'] = actual = persona.mi_plantilla_actual().rmupuesto
                    data['extras_actual'] = extrasactual = persona.horasextra_ingresos_mes()
                    data['rolpago_horasextras'] = persona.rolpago_horasextras()
                    data['rolpago_iess'] = persona.rolpago_iess()
                    data['rolpago_renta'] = persona.rolpago_renta()
                    data['porcentaje_seguro'] = PORCENTAJE_SEGURO
                    valorresta = 13
                    if not fechaingreso.year < hoy.year:
                        valorresta = int(valorresta - fechaingreso.month)
                    if persona.rolpago_ingresos_mes():
                        valorresta -= 1
                    proyectado = valorresta - hoy.month
                    data['rmu_proyectado'] = (proyectado * actual)
                    return render(request, "th_hojavida/addgasto.html", data)
                except Exception as ex:
                    pass

            elif action == 'delsolicitud':
                try:
                    data['title'] = u'Eliminar solicitud'
                    data['solicitud'] = SolicitudPublicacion.objects.get(pk=request.GET['id'])
                    return render(request, "th_hojavida/delsolicitud.html", data)
                except Exception as ex:
                    pass

            elif action == 'editsolicitud':
                try:
                    data['title'] = u'Editar Solicitud'
                    data['solicitud'] = solicitud = SolicitudPublicacion.objects.get(pk=request.GET['id'])
                    form = SolicitudPublicacionForm(initial={'tiposolicitud': solicitud.tiposolicitud,
                                                             'nombre': solicitud.nombre,
                                                             'motivo': solicitud.motivo,
                                                             'fecharecepcion': solicitud.fecharecepcion,
                                                             'fechaaprobacion': solicitud.fechaaprobacion,
                                                             'fechapublicacion': solicitud.fechapublicacion,
                                                             'volumen': solicitud.volumen,
                                                             'numero': solicitud.numero,
                                                             'paginas': solicitud.paginas,
                                                             'areaconocimiento': solicitud.areaconocimiento,
                                                             'subareaconocimiento': solicitud.subareaconocimiento,
                                                             'subareaespecificaconocimiento': solicitud.subareaespecificaconocimiento,
                                                             'evento': solicitud.evento,
                                                             'enlace': solicitud.enlace,
                                                             'revista': solicitud.revista,
                                                             'revista2': solicitud.revistainvestigacion,
                                                             'provieneproyecto': solicitud.provieneproyecto,
                                                             'lineainvestigacion': solicitud.lineainvestigacion,
                                                             'sublineainvestigacion': solicitud.sublineainvestigacion,
                                                             'base': solicitud.base,
                                                             'estadopublicacion': solicitud.estadopublicacion,
                                                             'tipoproyecto': solicitud.tipoproyecto,
                                                             'proyectointerno': solicitud.proyectointerno,
                                                             'proyectoexterno': solicitud.proyectoexterno,
                                                             'pais': solicitud.pais,
                                                             'ciudad': solicitud.ciudad,
                                                             'comitecientifico': solicitud.comitecientifico,
                                                             'estadopublicacionponencia': solicitud.estadopublicacion,
                                                             'fechapublicacionponencia': solicitud.fechapublicacion
                                                             })

                    form.editar(solicitud)

                    form2 = RevistaInvestigacionForm()

                    data['obligarchivo'] = 'N' if solicitud.estadopublicacion == 1 else 'S'

                    data['integrantescomite'] = solicitud.integrantecomite if solicitud.comitecientifico else ''
                    data['form'] = form
                    data['form2'] = form2
                    return render(request, "th_hojavida/editsolicitud.html", data)
                except Exception as ex:
                    pass

            elif action == 'editsolicitudlibro':
                try:
                    data['title'] = u'Editar Solicitud de Libro'
                    data['solicitud'] = solicitud = SolicitudPublicacion.objects.get(pk=request.GET['id'])
                    form = SolicitudPublicacionLibroForm(initial={'tiposolicitud': solicitud.tiposolicitud,
                                                                  'titulo': solicitud.nombre,
                                                                  'motivo': solicitud.motivo,
                                                                  'fechapublicacion': solicitud.fechapublicacion,
                                                                  'areaconocimiento': solicitud.areaconocimiento,
                                                                  'subareaconocimiento': solicitud.subareaconocimiento,
                                                                  'subareaespecificaconocimiento': solicitud.subareaespecificaconocimiento
                                                                  })

                    form.editar(solicitud)
                    data['form'] = form
                    return render(request, "th_hojavida/editsolicitudlibro.html", data)
                except Exception as ex:
                    pass

            elif action == 'editsolicitudcapitulo':
                try:
                    data['title'] = u'Editar Solicitud de capítulo de Libro'
                    data['solicitud'] = solicitud = SolicitudPublicacion.objects.get(pk=request.GET['id'])

                    form = CapituloLibroHojaVidaForm(initial={
                        'titulocapitulo': solicitud.nombre,
                        'resumen': solicitud.motivo,
                        'titulolibro': solicitud.evento,
                        'codigoisbn': solicitud.codigoisbn,
                        'paginas': solicitud.paginas,
                        'editorcompilador': solicitud.editorcompilador,
                        'fechapublicacion': solicitud.fechapublicacion,
                        'filiacion': solicitud.filiacion,
                        'areaconocimiento': solicitud.areaconocimiento,
                        'areaconocimiento': solicitud.areaconocimiento,
                        'subareaconocimiento': solicitud.subareaconocimiento,
                        'subareaespecificaconocimiento': solicitud.subareaespecificaconocimiento
                    })

                    form.editar(solicitud)
                    data['form'] = form
                    return render(request, "th_hojavida/editsolicitudcapitulo.html", data)
                except Exception as ex:
                    pass

            elif action == 'deleventoaprobacion':
                try:
                    data['title'] = u'Eliminar solicitud de evento'
                    data['cabecera'] = CapCabeceraSolicitud.objects.get(pk=int(request.GET['id']))
                    return render(request, "th_hojavida/deleventoaprobacion.html", data)
                except Exception as ex:
                    pass

            elif action == 'deleventoaprobacion_docente':
                try:
                    data['title'] = u'Eliminar solicitud de evento'
                    data['cabecera'] = CapCabeceraSolicitudDocente.objects.get(pk=int(request.GET['id']))
                    return render(request, "th_hojavida/deleventoaprobacion_docente.html", data)
                except Exception as ex:
                    pass

            elif action == 'pdf':
                try:
                    data['title'] = u'Adicionar Programa'
                    adduserdata(request, data)
                    persona = Persona.objects.get(pk=int(request.GET['id']))
                    # persona = request.session['persona']
                    profesor = persona.profesor().id
                    data['datospersona'] = persona
                    data['profesor'] = profesor
                    if ResponsableEvaluacion.objects.filter(status=True, activo=True).exists():
                        data['responsable'] = ResponsableEvaluacion.objects.filter(status=True, activo=True)[0]
                    data['periodoslect'] = MigracionEvaluacionDocente.objects.values('idperiodo', 'tipoeval').filter(
                        idprofesor=profesor, idperiodo=request.GET['idperiodo']).distinct()
                    data['detalleevaluacion'] = migra = MigracionEvaluacionDocente.objects.filter(idprofesor=profesor,
                                                                                                  idperiodo=request.GET[
                                                                                                      'idperiodo']).order_by(
                        'tipoeval', 'idperiodo', 'carrera', 'semestre', 'materia')
                    data['detalleevalconmodulo'] = migra.filter(modulo=1)
                    data['moduloevalcuatro'] = migra.filter(modulo=1, tipoeval=4)[0] if migra.filter(modulo=1,
                                                                                                     tipoeval=4).exists() else {}
                    data['detalleevalsinmodulo'] = migra.filter(modulo=0)
                    data['sinmoduloevalcuatro'] = migra.filter(modulo=0, tipoeval=4)[0] if migra.filter(modulo=0,
                                                                                                        tipoeval=4).exists() else {}
                    data['promperiodosinmodulo'] = promfinalc = round(
                        null_to_numeric(migra.filter(modulo=0).aggregate(prom=Avg('promedioasignatura'))['prom']), 2)
                    data['promperiodoconmodulo'] = promfinal = round(
                        null_to_numeric(migra.filter(modulo=1).aggregate(prom=Avg('promedioasignatura'))['prom']), 2)
                    if promfinal:
                        notaf = str(request.GET['idperiodo']) + "-" + persona.cedula + "-" + str(promfinal)
                    else:
                        notaf = str(request.GET['idperiodo']) + "-" + persona.cedula + "-" + str(promfinalc)
                    data['nomperiodo'] = request.GET['nomperiodo']
                    data['tipoev'] = request.GET['tipoev']
                    qrname = 'qrce_evam_' + request.GET['idperiodo'] + persona.cedula
                    folder = os.path.join(os.path.join(SITE_STORAGE, 'media', 'qrcode', 'evaluaciondocente', 'qr'))
                    # folder = '/mnt/nfs/home/storage/media/qrcode/evaluaciondocente/'
                    # folder = '/mnt/nfs/home/storage/media/qrcode/evaluaciondocente/'
                    rutapdf = folder + qrname + '.pdf'
                    rutaimg = folder + qrname + '.png'
                    if os.path.isfile(rutapdf):
                        os.remove(rutaimg)
                        os.remove(rutapdf)
                    url = pyqrcode.create('http://sga.unemi.edu.ec//media/qrcode/evaluaciondocente/' + qrname + '.pdf')
                    imageqr = url.png(folder + qrname + '.png', 16, '#000000')
                    imagebarcode = code128.image(notaf).save(folder + qrname + "_bar.png")
                    data['qrname'] = 'qr' + qrname
                    data['fechactual'] = datetime.now().strftime("%d") + '/' + datetime.now().strftime(
                        "%m") + '/' + datetime.now().strftime("%y") + ' ' + datetime.now().strftime("%H:%M")
                    return conviert_html_to_pdfsave('pro_certificados/certificado_porperiodo.html',
                                                    {
                                                        'pagesize': 'A4',
                                                        'listadoevaluacion': data,
                                                    }, qrname + '.pdf'
                                                    )
                except Exception as ex:
                    pass

            elif action == 'pdfmodelo2015':
                try:
                    data['title'] = u'Adicionar Programa'
                    adduserdata(request, data)
                    persona = Persona.objects.get(pk=int(request.GET['id']))
                    profesor = persona.profesor().id
                    data['datospersona'] = persona
                    data['profesor'] = profesor
                    if ResponsableEvaluacion.objects.filter(status=True, activo=True).exists():
                        data['responsable'] = ResponsableEvaluacion.objects.filter(status=True, activo=True)[0]
                    data['nomperiodo'] = request.GET['nomperiodo']
                    data['resultados'] = notaporcentaje = ResumenParcialEvaluacionIntegral.objects.filter(
                        profesor=profesor, proceso=request.GET['idperiodo']).order_by(
                        'materia__asignaturamalla__malla__carrera__id', 'materia__asignaturamalla__nivelmalla__id')
                    data['porcentaje'] = round(
                        null_to_numeric(notaporcentaje.aggregate(prom=Avg('totalmateriadocencia'))['prom']), 2)
                    data['fechactual'] = datetime.now().strftime("%d") + '/' + datetime.now().strftime(
                        "%m") + '/' + datetime.now().strftime("%y") + ' ' + datetime.now().strftime("%H:%M")
                    notaporcen = "0" + str(request.GET['idperiodo']) + "-" + persona.cedula + "-" + str(
                        data['porcentaje'])
                    qrname = 'qrce_2015_0' + request.GET['idperiodo'] + persona.cedula
                    # folder = SITE_STORAGE + 'media/qrcode/evaluaciondocente/'
                    folder = os.path.join(os.path.join(SITE_STORAGE, 'media', 'qrcode', 'evaluaciondocente', 'qr'))
                    rutapdf = folder + qrname + '.pdf'
                    rutaimg = folder + qrname + '.png'
                    if os.path.isfile(rutapdf):
                        os.remove(rutaimg)
                        os.remove(rutapdf)
                    url = pyqrcode.create('http://sga.unemi.edu.ec//media/qrcode/evaluaciondocente/' + qrname + '.pdf')
                    imageqr = url.png(folder + qrname + '.png', 16, '#000000')
                    imagebarcode = code128.image(notaporcen).save(folder + qrname + "_bar.png")
                    data['qrname'] = 'qr' + qrname
                    return conviert_html_to_pdfsave(
                        'pro_certificados/pdfqrce_2015.html',
                        {
                            'pagesize': 'A4',
                            'listadoevaluacion': data,
                        }, qrname + '.pdf'
                    )
                except Exception as ex:
                    pass

            elif action == 'addarchivo':
                try:
                    # puede_realizar_accion(request, 'sga.puede_modificar_inscripciones')
                    if perfilprincipal.es_estudiante:
                        data['title'] = u'Adicionar Documento'
                        data['inscripcion'] = Inscripcion.objects.get(pk=request.GET['id'])
                        # puede_modificar_inscripcion(request, inscripcion)
                        form = DocumentoInscripcionForm()
                        form.ocultar_nombre()
                        data['form'] = form
                        return render(request, "th_hojavida/addarchivo.html", data)
                except Exception as ex:
                    pass

            elif action == 'editarchivo':
                try:
                    # puede_realizar_accion(request, 'sga.puede_modificar_inscripciones')
                    if perfilprincipal.es_estudiante:
                        data['title'] = u'Editar Documento'
                        data['archivo'] = archivo = Archivo.objects.get(pk=int(request.GET['idar']))
                        data['inscripcion'] = perfilprincipal.inscripcion
                        # puede_modificar_inscripcion(request, inscripcion)
                        form = DocumentoInscripcionForm(initial={'nombre': archivo.nombre, 'tipo': archivo.tipo})
                        form.ocultar_nombre()
                        data['form'] = form
                        return render(request, "th_hojavida/editarchivo.html", data)
                except Exception as ex:
                    pass

            elif action == 'delarchivo':
                try:
                    # puede_realizar_accion(request, 'sga.puede_modificar_inscripciones')
                    if perfilprincipal.es_estudiante:
                        data['title'] = u'Eliminar Documento Personal'
                        data['archivo'] = Archivo.objects.get(pk=int(request.GET['idar']))
                        # puede_modificar_inscripcion(request, inscripcion)
                        return render(request, "th_hojavida/delarchivo.html", data)
                except Exception as ex:
                    pass

            elif action == 'editarticulos':
                try:
                    data['title'] = u'Editar articulo'
                    data['articulos'] = articulos = ArticuloInvestigacion.objects.get(pk=request.GET['id'])
                    form = ArticuloInvestigacionForm(initial={'areaconocimiento': articulos.areaconocimiento,
                                                              'subareaconocimiento': articulos.subareaconocimiento,
                                                              'subareaespecificaconocimiento': articulos.subareaespecificaconocimiento})
                    form.editar(articulos)
                    data['form'] = form
                    return render(request, "th_hojavida/editarticulos.html", data)
                except Exception as ex:
                    pass

            elif action == 'pdfmodeloactual':
                try:
                    import statistics
                    data['title'] = u'Adicionar Programa'
                    adduserdata(request, data)
                    profesor = Profesor.objects.get(pk=int(request.GET['profesor']))
                    data['departamento'] = departamento = Departamento.objects.get(pk=128)
                    respuestas = []
                    promediovirtual = 0
                    data['procesoperiodo'] = ProcesoEvaluativoAcreditacion.objects.get(periodo=request.GET['idperiodo'],
                                                                                       status=True)
                    if RubricaPreguntas.objects.filter(rubrica__tipo_criterio=1, rubrica__informativa=False,
                                                       rubrica__para_nivelacionvirtual=True,
                                                       detallerespuestarubrica__respuestarubrica__respuestaevaluacion__tipoinstrumento=1,
                                                       detallerespuestarubrica__respuestarubrica__respuestaevaluacion__profesor=profesor,
                                                       detallerespuestarubrica__respuestarubrica__respuestaevaluacion__proceso__periodo=
                                                       request.GET['idperiodo']).distinct().exists():
                        for prome in RubricaPreguntas.objects.values_list('id').filter(rubrica__tipo_criterio=1,
                                                                                       rubrica__informativa=False,
                                                                                       rubrica__para_nivelacionvirtual=True,
                                                                                       detallerespuestarubrica__respuestarubrica__respuestaevaluacion__tipoinstrumento=1,
                                                                                       detallerespuestarubrica__respuestarubrica__respuestaevaluacion__profesor=profesor,
                                                                                       detallerespuestarubrica__respuestarubrica__respuestaevaluacion__proceso__periodo=
                                                                                       request.GET[
                                                                                           'idperiodo']).annotate(
                            prom=Avg('detallerespuestarubrica__valor')):
                            respuestas.append(null_to_decimal(prome[1], 2))
                        promedio = statistics.mean(respuestas) if respuestas else 0
                        promediovirtual = null_to_decimal(promedio, 2) if promedio else 0
                    data['promediovirtual'] = promediovirtual
                    promedionovirtual = 0
                    respuestas2 = []
                    if RubricaPreguntas.objects.filter(rubrica__tipo_criterio=1, rubrica__informativa=False,
                                                       rubrica__para_nivelacionvirtual=False,
                                                       detallerespuestarubrica__respuestarubrica__respuestaevaluacion__tipoinstrumento=1,
                                                       detallerespuestarubrica__respuestarubrica__respuestaevaluacion__profesor=profesor,
                                                       detallerespuestarubrica__respuestarubrica__respuestaevaluacion__proceso__periodo=
                                                       request.GET['idperiodo']).distinct().exists():
                        for prome in RubricaPreguntas.objects.values_list('id').filter(rubrica__tipo_criterio=1,
                                                                                       rubrica__informativa=False,
                                                                                       rubrica__para_nivelacionvirtual=False,
                                                                                       detallerespuestarubrica__respuestarubrica__respuestaevaluacion__tipoinstrumento=1,
                                                                                       detallerespuestarubrica__respuestarubrica__respuestaevaluacion__profesor=profesor,
                                                                                       detallerespuestarubrica__respuestarubrica__respuestaevaluacion__proceso__periodo=
                                                                                       request.GET[
                                                                                           'idperiodo']).annotate(
                            prom=Avg('detallerespuestarubrica__valor')):
                            respuestas2.append(null_to_decimal(prome[1], 2))
                        promedio = statistics.mean(respuestas2) if respuestas2 else 0
                        promedionovirtual = null_to_decimal(promedio, 2) if promedio else 0
                    data['promedionovirtual'] = promedionovirtual
                    data['datospersona'] = profesor.persona
                    data['profesor'] = profesor
                    if ResponsableEvaluacion.objects.filter(status=True, activo=True).exists():
                        data['responsable'] = ResponsableEvaluacion.objects.filter(status=True, activo=True)[0]
                    data['nomperiodo'] = request.GET['nomperiodo']
                    data['resultados'] = porcentaje = ResumenFinalEvaluacionAcreditacion.objects.get(
                        distributivo__profesor=profesor, distributivo__periodo=request.GET['idperiodo'])
                    data['porcentaje'] = notaporcentaje = round(((porcentaje.resultado_total * 100) / 5), 2)
                    data['fechactual'] = datetime.now().strftime("%d") + '/' + datetime.now().strftime(
                        "%m") + '/' + datetime.now().strftime("%y") + ' ' + datetime.now().strftime("%H:%M")
                    notaporcen = str(request.GET['idperiodo']) + "-" + profesor.persona.cedula + "-" + str(
                        notaporcentaje)
                    qrname = 'qrce_mied_' + request.GET['idperiodo'] + profesor.persona.cedula
                    # folder = SITE_STORAGE + 'media/qrcode/evaluaciondocente/'
                    folder = os.path.join(os.path.join(SITE_STORAGE, 'media', 'qrcode', 'evaluaciondocente', 'qr'))
                    # folder = os.path.join(SITE_STORAGE, 'media', 'qrcode', 'evaluaciondocente')
                    rutapdf = folder + qrname + '.pdf'
                    rutaimg = folder + qrname + '.png'
                    if os.path.isfile(rutapdf):
                        os.remove(rutaimg)
                        os.remove(rutapdf)
                    url = pyqrcode.create('http://sga.unemi.edu.ec/media/qrcode/evaluaciondocente/' + qrname + '.pdf')
                    imageqr = url.png(folder + qrname + '.png', 16, '#000000')
                    imagebarcode = code128.image(notaporcen).save(folder + qrname + "_bar.png")
                    data['qrname'] = 'qr' + qrname
                    data['url_path'] = 'http://127.0.0.1:8000'
                    if not DEBUG:
                        data['url_path'] = 'https://sga.unemi.edu.ec'
                    return conviert_html_to_pdfsave(
                        'pro_certificados/pdfqrce_modeloactual.html',
                        {
                            'pagesize': 'A4',
                            'listadoevaluacion': data,
                        }, qrname + '.pdf'
                    )
                except Exception as ex:
                    pass

            elif action == 'delcertificado':
                try:
                    data['title'] = u'Eliminar Certificación'
                    data['certificado'] = CertificacionPersona.objects.get(pk=request.GET['id'])
                    return render(request, "th_hojavida/delcertificado.html", data)
                except Exception as ex:
                    pass

            elif action == 'editotrocertificado':
                try:
                    data['title'] = u'Editar Certificado Idioma'
                    data['certificado'] = certificado = CertificadoIdioma.objects.get(pk=request.GET['id'])

                    data['form'] = CertificadoIdiomaForm(initial={'idioma': certificado.idioma,

                                                                  'institucion': certificado.institucioncerti,
                                                                  'otrainstitucion': certificado.otrainstitucion,
                                                                  'validainst': certificado.validainst,
                                                                  'fecha': certificado.fechacerti,
                                                                  'nivel': certificado.nivelsuficencia})
                    return render(request, "th_hojavida/editotrocertificado.html", data)
                except Exception as ex:
                    pass

            elif action == 'logmarcadas':
                try:
                    data['title'] = u'LOG de Marcadas'
                    # data['distributivo'] = distributivo = DistributivoPersona.objects.get(pk=int(encrypt(request.GET['id'])))
                    data['personaadminis'] = persona
                    data['anios'] = persona.lista_anios_trabajados_log()
                    data['jornadas'] = persona.historialjornadatrabajador_set.all()
                    puede_crear_marcada = False
                    try:
                        puede_crear_marcada = puede_realizar_accion_is_superuser(request, 'sagest.puede_crear_marcada')
                    except:
                        puede_crear_marcada = False
                    data['puede_crear_marcada'] = puede_crear_marcada
                    data['destino'] = 'th_hojavida'
                    data['pued_modificar'] = 1
                    # if not persona.id == persona.id:
                    #     raise NameError('Error')
                    data['hora'] = str(datetime.now().time())[0:5]
                    return render(request, "th_marcadas/logmarcadas.html", data)
                except Exception as ex:
                    pass

            elif action == 'solicitudjusti':
                try:
                    data['title'] = u'Solicitud justificación de marcadas'
                    data['form'] = SolicitudJustificacionMarcadaForm()
                    # template = get_template("th_hojavida/modal/solicitudjustificcion.html")
                    return render(request, "th_hojavida/modal/solicitudjustificcion.html", data)
                except Exception as ex:
                    pass

            elif action == 'horasmarcadasfec':
                try:
                    lista = []
                    fecha = convertir_fecha_invertida(request.GET.get('fecha'))
                    if not LogDia.objects.filter(status=True, persona=persona, fecha=fecha).exists():
                        raise NameError('No existe marcadas ese día, intente con otra fecha.')
                    logdia = LogDia.objects.get(status=True, persona=persona, fecha=fecha)
                    logmarcada = logdia.logmarcada_set.filter(status=True)
                    for mar in logmarcada:
                        lista.append([mar.pk, '%s' % datetime.strftime(mar.time, '%H:%M')])
                    return JsonResponse({'result': True, 'marcada': lista})
                except Exception as ex:
                    return JsonResponse(
                        {'result': False, 'mensaje': 'Error al procesar los datos!. %s' % (ex.__str__())})

            elif action == 'justificacionmarcada':
                try:
                    if 'id' in request.GET:
                        logdia = LogDia.objects.get(pk=encrypt(request.GET['id']))
                        data['solicitud'] = soli = logdia.solicitud()
                        data['cabecera'] = his = soli[0].solicitud
                        data['historial'] = his.historialsolicitudjustificacionmarcada_set.filter(status=True).all()
                    elif 'pk' in request.GET:
                        if request.user.has_perm('sagest.puede_aprobar_justificacion_marcada_director'):
                            data['es_director_th'] = True
                        elif request.user.has_perm('sagest.puede_cerrar_justificacion_marcada_analista'):
                            data['es_analista_th'] = True
                        id = int(encrypt(request.GET['pk']))
                        data['cabecera'] = soli = SolicitudJustificacionMarcada.objects.get(pk=id, status=True,
                                                                                            solicita=int(
                                                                                                request.GET['persona']))
                        data['solicitud'] = soli.detallesolicitudjustificacionmarcada_set.filter(status=True).all()
                        data['historial'] = soli.historialsolicitudjustificacionmarcada_set.filter(status=True).all()
                        data['aprueba'] = True
                        data['fecha'] = datetime.now()
                    template = get_template("th_hojavida/modal/detallejustificacion.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content})
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})



            elif action == 'buscarpersonas':
                try:
                    resp = filtro_persona_select(request)
                    return HttpResponse(json.dumps({'status': True, 'results': resp}))
                except Exception as ex:
                    pass

            elif action == 'historialjustificacion':
                try:
                    id = int(encrypt(request.GET['pk']))
                    data['cabecera'] = soli = SolicitudJustificacionMarcada.objects.get(pk=id, status=True,
                                                                                        solicita=int(
                                                                                            request.GET['persona']))
                    data['solicitud'] = soli.historialsolicitudjustificacionmarcada_set.filter(status=True).all()
                    template = get_template("th_hojavida/modal/historialjustificacion.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content})
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos"})

            elif action == 'addcabsolicitud':
                try:
                    data['title'] = u'Confirmar solicitud de evento'
                    data['evento'] = CapEventoPeriodo.objects.get(pk=int(encrypt(request.GET['id'])))
                    return render(request, "th_hojavida/addcabsolicitud.html", data)
                except:
                    pass

            elif action == 'addcabsolicitud_docente':
                try:
                    data['title'] = u'Confirmar solicitud de evento'
                    data['evento'] = CapEventoPeriodoDocente.objects.get(pk=int(encrypt(request.GET['id'])))
                    return render(request, "th_hojavida/addcabsolicitud_docente.html", data)
                except:
                    pass

            elif action == 'preferencia':
                try:
                    data['title'] = u'Seleccione los parametros de preferencias de asignaturas'
                    data['admisionperiodo'] = periodo = request.session['periodoadmision']
                    data['persona'] = persona
                    if persona.es_administrador() or persona.es_profesor():
                        fecha = datetime.now().date()
                        if periodo.preferenciaadmisionfechatope:
                            if (periodo.preferenciaadmisionfechatope >= fecha):
                                data['accesopreferencia'] = True
                            else:
                                data['accesopreferencia'] = False
                        else:
                            data['accesopreferencia'] = False
                    else:
                        data['accesopreferencia'] = False
                    cordinaciones = Coordinacion.objects.filter(pk=9, carrera__in=Materia.objects.values_list(
                        'asignaturamalla__malla__carrera').filter(nivel__periodo=periodo).distinct()).distinct()
                    data['cordinaciones'] = cordinaciones
                    return render(request, "th_hojavida/preferencias.html", data)
                except Exception as ex:
                    pass

            elif action == 'participantepublicacion':
                try:
                    data['title'] = u'Participantes de la Publicación'
                    data['solicitud'] = solicitud = SolicitudPublicacion.objects.get(pk=request.GET['id'])
                    data['tipoparinstitucion'] = TIPO_PARTICIPANTE_INSTITUCION
                    data['participantes'] = ParticipanteSolicitudPublicacion.objects.filter(status=True,
                                                                                            solicitud=solicitud)
                    return render(request, "th_hojavida/participantepublicacion.html", data)
                except Exception as ex:
                    pass

            elif action == 'addparticipantedocente':
                try:
                    data['title'] = u'Participante Docente'
                    data['form'] = ParticipanteProfesorArticuloForm
                    data['id'] = request.GET['idsolicitud']
                    data['titulo'] = SolicitudPublicacion.objects.get(pk=request.GET['idsolicitud']).nombre
                    return render(request, "th_hojavida/addparticipantedocente.html", data)
                except Exception as ex:
                    pass

            elif action == 'addparticipanteadministrativo':
                try:
                    data['title'] = u'Participante Administrativo'
                    data['form'] = ParticipanteAdministrativoArticuloForm
                    data['id'] = request.GET['idsolicitud']
                    data['titulo'] = SolicitudPublicacion.objects.get(pk=request.GET['idsolicitud']).nombre
                    return render(request, "th_hojavida/addparticipanteadministrativo.html", data)
                except Exception as ex:
                    pass

            elif action == 'addparticipanteinscripcion':
                try:
                    data['title'] = u'Participante Estudiante'
                    data['form'] = ParticipanteInscripcionArticuloForm
                    data['id'] = request.GET['idsolicitud']
                    data['titulo'] = SolicitudPublicacion.objects.get(pk=request.GET['idsolicitud']).nombre
                    return render(request, "th_hojavida/addparticipanteinscripcion.html", data)
                except Exception as ex:
                    pass

            elif action == 'deleteparticipantepublicacion':
                try:
                    data['title'] = u'Eliminar Participante'
                    tipo = request.GET['tipo']
                    data['participante'] = participante = ParticipanteSolicitudPublicacion.objects.get(
                        pk=request.GET['id'])
                    if tipo == '1':
                        data['nombres'] = participante.profesor.persona.nombre_completo()
                    if tipo == '3':
                        data['nombres'] = participante.administrativo.persona.nombre_completo()
                    if tipo == '4':
                        data['nombres'] = participante.inscripcion.persona.nombre_completo()
                    return render(request, "th_hojavida/deleteparticipantepublicacion.html", data)
                except Exception as ex:
                    pass

            elif action == 'informacion':
                try:
                    data['formulario'] = formulario = Formulario107.objects.get(pk=request.GET['id'])
                    data['title'] = u'%s' % formulario.persona.nombre_completo_inverso()
                    template = get_template("th_hojavida/informacion.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content})
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

            elif action == 'addbitacora':
                try:
                    hoy = datetime.now()
                    data['title'] = u'Adicionar Bitácora de actividad'
                    periodo = request.session['periodo']
                    form = BitacoraForm()
                    ver_tipoactividad = False
                    if persona.contratodip_set.filter(status=True).exclude(estado=5).exists():
                        data['tienecontrato'] = True
                        fecha_comienza = variable_valor("VALIDAR_FECHA_LIMITE_POSGRADO")
                        if hoy.date() >= fecha_comienza:
                            data['contrato_posgrado'] = True
                        contratodip = persona.contratodip_set.filter(status=True).exclude(estado=5).order_by('fechafin').last()
                        data['contratodip'] = contratodip
                        form.ocultarcampos_titulo()
                        actividades = contratodip.cargo.actividadescontratoperfil_set.filter(status=True)
                        actextra = contratodip.actividadescontratoperfil_set.filter(status=True)
                        acti = actividades.values_list('actividad_id', flat=True) | actextra.values_list('actividad_id', flat=True)
                        form.fields['actividades'].queryset = ActividadesPerfil.objects.filter(status=True, pk__in=acti)
                        if Departamento.objects.filter(pk=93).exists():
                            departamento = Departamento.objects.get(pk=93)
                            if persona.id in departamento.mis_integrantes().values_list('id', flat=True):
                                ver_tipoactividad = True

                        if CoordinadorCarrera.objects.values('id').filter(Q(persona=persona, status=True) & (
                                Q(periodo__clasificacion=2) | Q(periodo__tipo=3))).exists():
                            data["es_coordinador"] = True
                            form.fields['departamento_requiriente'].initial = 162
                        else:
                            form.ocultarcampos_horafin()
                    else:
                        form.ocultarcampos_actividades()
                    if persona.grupos().values("id").filter(pk=49).exists():
                        ver_tipoactividad = True

                    form.ocultarcampos_tipoactividad() if not ver_tipoactividad else None
                    data['form'] = form
                    return render(request, "th_hojavida/addbitacora.html", data)
                except Exception as ex:
                    pass

            elif action == 'validarhorario':
                try:
                    fi, hi, hf = request.GET['fi'], request.GET['hi'], request.GET['hf']

                    mensaje = u"La %s ingresada es incorrecta. Por favor revisar."
                    if not fi:
                        return {"result": False, 'mensaje': mensaje % 'fecha'}

                    if not hi:
                        return {"result": False, 'mensaje': mensaje % 'hora de inicio'}

                    if not hf:
                        return {"result": False, 'mensaje': mensaje % u'hora de finalización'}

                    fecha = convertir_fecha_invertida(fi)
                    listaturnosocupados = validar_choque_horario_actividad_gestion(persona, fecha, hi, hf,
                                                                                   request.session['periodo'])
                    if listaturnosocupados[0]:
                        data['choqueturno'] = listaturnosocupados[0]
                        data['listaturnosocupados'] = listaturnosocupados[1]
                        data['info'] = {'fecha': fecha, 'hi': hi, 'hf': hf}
                        template = get_template("th_hojavida/modal/listaturnosdisponibles.html")
                        return JsonResponse({"result": False, 'data': template.render(data)})
                    else:
                        return JsonResponse({"result": True, 'mensaje': 'Horario valido.'})

                except Exception as ex:
                    return JsonResponse({"result": False, 'mensaje': 'Error de conexión'})

            elif action == 'editbitacora':
                try:
                    data['title'] = u'Editar bitácora'
                    data['bitacora'] = bitacora = BitacoraActividadDiaria.objects.get(
                        pk=int(encrypt(request.GET['id'])))
                    form = BitacoraForm(initial=model_to_dict(bitacora))
                    form.fields['hora'].initial = bitacora.fecha.time() if bitacora.fecha else None
                    form.fields['horafin'].initial = bitacora.fechafin.time() if bitacora.fechafin else None
                    hoy = datetime.now()
                    if persona.contratodip_set.filter(status=True).exclude(estado=5).exists():
                        data['tienecontrato'] = True
                        # form.bloquear_fecha_hora()
                        fecha_comienza = variable_valor("VALIDAR_FECHA_LIMITE_POSGRADO")
                        if hoy.date() >= fecha_comienza:
                            data['contrato_posgrado'] = True
                        contratodip = persona.contratodip_set.filter(status=True).exclude(estado=5).order_by('fechafin').last()
                        form.ocultarcampos_titulo()
                        data['contratodip'] = contratodip
                        # actividades = contratodip.cargo.actividades.filter(status=True).all()
                        # actextra = contratodip.actividadesextra.filter(status=True).all()
                        # form.fields['actividades'].queryset = actividades|actextra
                        actividades = contratodip.cargo.actividadescontratoperfil_set.filter(status=True)
                        actextra = contratodip.actividadescontratoperfil_set.filter(status=True)
                        acti = actividades.values_list('actividad_id', flat=True) | actextra.values_list('actividad_id',
                                                                                                         flat=True)
                        form.fields['actividades'].queryset = ActividadesPerfil.objects.filter(status=True, pk__in=acti)
                        form.fields['actividades'].initial = [bitacora.actividades]
                        if Departamento.objects.filter(pk=93).exists():
                            departamento = Departamento.objects.get(pk=93)
                            if not persona.id in departamento.mis_integrantes().values_list('id', flat=True):
                                if not persona.grupos().values("id").filter(pk=49).exists():
                                    form.ocultarcampos_tipoactividad()

                        if not CoordinadorCarrera.objects.values('id').filter(persona=persona, status=True).exists():
                            form.ocultarcampos_horafin()
                    else:
                        if not persona.grupos().values("id").filter(pk=49).exists():
                            form.ocultarcampos_actividades()

                    data['form'] = form
                    data['action'] = 'editbitacora'
                    return render(request, "th_hojavida/addbitacora.html", data)
                except Exception as ex:
                    pass

            elif action == 'addbitacoraposgrado':
                try:
                    data['title'] = u'Adicionar Bitácora de actividad'
                    form = BitacoraForm()
                    contratodip = persona.contratodip_set.filter(status=True).last()
                    form.ocultarcampos_titulo()
                    actividades = contratodip.cargo.actividades.filter(status=True).all()
                    actextra = contratodip.actividadesextra.filter(status=True).all()
                    form.fields['actividades'].queryset = actividades | actextra
                    data['form'] = form
                    data['action'] = 'addbitacoraposgrado'
                    return render(request, "th_hojavida/addbitacora.html", data)
                except Exception as ex:
                    pass

            elif action == 'editbitacoraposgrado':
                try:
                    data['title'] = u'Editar bitácora'
                    data['bitacora'] = bitacora = BitacoraActividadDiaria.objects.get(
                        pk=int(encrypt(request.GET['id'])))

                    initial = model_to_dict(bitacora)
                    form = BitacoraForm(initial=initial)
                    contratodip = persona.contratodip_set.filter(status=True).last()
                    form.ocultarcampos_titulo()
                    actividades = contratodip.cargo.actividades.filter(status=True).all()
                    actextra = contratodip.actividadesextra.filter(status=True).all()
                    form.fields['actividades'].queryset = actividades | actextra
                    data['form'] = form
                    data['action'] = 'editbitacoraposgrado'
                    return render(request, "th_hojavida/editbitacora.html", data)
                except Exception as ex:
                    pass

            elif action == 'delbitacora':
                try:
                    data['title'] = u'Eliminar bitácora'
                    data['bitacora'] = BitacoraActividadDiaria.objects.get(pk=int(encrypt(request.GET['id'])))
                    return render(request, "th_hojavida/delbitacora.html", data)
                except Exception as ex:
                    pass

            elif action == 'addvacunacvd19aspirante':
                try:
                    data['form2'] = VacunacionCovidAdmisionForm()
                    template = get_template("th_hojavida/modal/formvacunacionaspirante.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'addarchivocontratofirmado':
                try:

                    data['filtro'] = filtro = PersonaContratos.objects.get(pk=int(request.GET['id']))
                    form = ArchivoHistorialContratoForm()
                    data['estado'] = int(request.GET['tipo'])
                    data['form2'] = form
                    template = get_template("th_hojavida/modal/formarchivocontrato.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'adddosisvacunaaspirante':
                try:
                    data['filtro'] = filtro = VacunaCovid.objects.get(pk=int(request.GET['id']))
                    data['formmodal'] = VacunacionCovidAdmisionForm()
                    template = get_template("th_hojavida/modal/formdosisaspirante.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'addcertificadovacunacovidaspirante':
                try:
                    data['filtro'] = filtro = VacunaCovid.objects.get(pk=int(request.GET['id']))
                    data['form2'] = VacunacionCovidEvidenciaAdmisionForm()
                    template = get_template("th_hojavida/modal/formvacunacionevaspirante.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'reporte_bitacora_excel':
                try:
                    fechadesde = convertir_fecha_invertida(request.GET['fecha_desde'])
                    fechahasta = convertir_fecha_invertida(request.GET['fecha_hasta'])
                    __author__ = 'Unemi'
                    title = easyxf(
                        'font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                    title2 = easyxf(
                        'font: name Times New Roman, color-index black, bold on , height 250; alignment: horiz centre')
                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('Hoja1')
                    bitacoras = BitacoraActividadDiaria.objects.filter(status=True, fecha__date__gte=fechadesde,
                                                                       fecha__date__lte=fechahasta, persona=persona)
                    ws.write_merge(0, 0, 0, 5, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                    ws.write_merge(1, 1, 0, 5, 'BITÁCORA DE ACTIVIDADES DIARIAS', title2)
                    ws.write_merge(2, 2, 0, 5, f'{persona}', title2)
                    ws.write_merge(3, 3, 0, 5, f'DESDE  {fechadesde} HASTA {fechahasta}', title2)
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=bitacora' + random.randint(1,
                                                                                                       10000).__str__() + '.xls'
                    columns = [
                        (u"DEPARTAMENTO", 6000),
                        (u"FECHA", 6000),
                        (u"TITULO", 6000),
                        (u"DESCRIPCIÓN", 15000),
                        (u"LINK", 6000),
                        (u"TIPO SISTEMA", 6000),
                    ]
                    row_num = 5
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], font_style)
                        ws.col(col_num).width = columns[col_num][1]

                    row_num = 6
                    for r in bitacoras:
                        campo1 = u"%s" % r.departamento.nombre if r.departamento else 'SIN DEPARTAMENTO'
                        campo2 = u"%s" % r.fecha
                        campo3 = u"%s" % r.titulo
                        campo4 = u"%s" % r.descripcion
                        campo5 = u"%s" % r.link
                        campo6 = u"%s" % r.get_tiposistema_display()
                        ws.write(row_num, 0, campo1, font_style2)
                        ws.write(row_num, 1, campo2, font_style2)
                        ws.write(row_num, 2, campo3, font_style2)
                        ws.write(row_num, 3, campo4, font_style2)
                        ws.write(row_num, 4, campo5, font_style2)
                        ws.write(row_num, 5, campo6, font_style2)
                        row_num += 1
                    wb.save(response)
                    return response
                except Exception as ex:
                    messages.error(request, str(ex))
                    # pass

            elif action == 'reporte_bitacora_pdf':
                try:
                    data['fechadesde'] = fechadesde = convertir_fecha(request.GET['fecha_desde'])
                    data['fechahasta'] = fechahasta = convertir_fecha(request.GET['fecha_hasta'])
                    data['bitacoras'] = BitacoraActividadDiaria.objects.filter(status=True, fecha__date__gte=fechadesde,
                                                                               fecha__date__lte=fechahasta,
                                                                               persona=persona)
                    return conviert_html_to_pdf(
                        'adm_requerimiento/reporte_bitacora_pdf.html',
                        {
                            'pagesize': 'a4 landscape',
                            'data': data,
                        }
                    )
                except Exception as ex:
                    pass

            elif action == 'editmarcada':
                try:
                    logmarcada = LogMarcada.objects.get(pk=request.GET['id'])
                    form = CambiarMarcadaForm(initial={'hora': logmarcada.time.time()})
                    data['form'] = form
                    data['id'] = request.GET['id']
                    template = get_template("th_marcadas/modal/formmarcada.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'detalle_contrato':
                try:
                    if 'id' in request.GET:
                        data['contrato'] = ContratoDip.objects.get(id=request.GET['id'],
                                                                   status=True)
                        template = get_template("adm_contratodip/detallecontrato.html")
                        return JsonResponse({"result": 'ok', 'data': template.render(data)})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

            elif action == 'viewreportes':
                try:
                    if 'idcon' in request.GET:
                        data['contrato'] = contratopersona = ContratoDip.objects.get(status=True,
                                                                                     id=encrypt(request.GET['idcon']))
                        data['historial'] = historial = HistorialPagoMes.objects.filter(status=True,
                                                                                        contrato=contratopersona).order_by(
                            'fecha_pago')
                        data['tecnico'] = InformeTecnico.objects.filter(status=True, contrato=contratopersona,
                                                                        historialpago=historial)
                        data['bitacora'] = InformeActividadJornada.objects.filter(status=True, contrato=contratopersona,
                                                                                  historialpago=historial)
                        data['title'] = 'Registro de Reportes'
                        template = get_template("th_hojavida/viewreportes.html")
                        return JsonResponse({"result": 'ok', 'data': template.render(data)})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

            elif action == 'consultaarchivos':
                try:
                    data['contrato'] = contratopersona = ContratoDip.objects.get(status=True,
                                                                                 id=encrypt(request.GET['idcon']))
                    data['historial'] = historial = HistorialPagoMes.objects.get(status=True, contrato=contratopersona,
                                                                                 id=encrypt(request.GET['idhist']))
                    data['actapago'] = ActaPago.objects.filter(status=True, contrato=contratopersona,
                                                               fecha_creacion__year=historial.fecha_pago.year,
                                                               mes=historial.fecha_pago.month).first()
                    data['tecnico'] = InformeTecnico.objects.filter(status=True, contrato=contratopersona,
                                                                    fecha_creacion__year=historial.fecha_pago.year,
                                                                    mes=historial.fecha_pago.month).first()
                    data['bitacora'] = InformeActividadJornada.objects.filter(status=True, contrato=contratopersona,
                                                                              fecha_creacion__year=historial.fecha_pago.year,
                                                                              mes=historial.fecha_pago.month).first()
                    data['memos'] = MemoActividadPosgrado.objects.filter(status=True, contrato=contratopersona,
                                                                         fecha_creacion__year=historial.fecha_pago.year,
                                                                         mes=historial.fecha_pago.month).first()
                    if ContratoDip.objects.filter(status=True, id=encrypt(request.GET['idcon']),
                                                  persona=persona).exists():
                        data['trabaja'] = True
                    template = get_template("adm_contratodip/archivospagos.html")
                    return JsonResponse({"result": 'ok', 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'bitacora':
                try:
                    ids = request.GET['id']
                    desde = request.GET['desde']
                    data['distributivo'] = contratopersona = ContratoDip.objects.filter(pk=int(ids))[0]
                    data['persona'] = personad = Persona.objects.get(pk=contratopersona.persona.id)
                    cumplehoras = TrabajadorDiaJornada.objects.filter(persona=contratopersona.persona, mes=int(desde),
                                                                      status=True)
                    if InformeActividadJornada.objects.values('id').filter(mes=int(desde), status=True,
                                                                           contrato=contratopersona).exists():
                        return JsonResponse({'result': False, 'mensaje': 'Ya existe el registro con el mes de %s' % (
                        MESES_CHOICES[int(desde) - 1][1])})
                    secinforme = contratopersona.secuencia_informe()
                    if not (
                    HistorialPagoMes.objects.values('id').filter(status=True, contrato_id=int(request.GET['id']),
                                                                 fecha_pago__month=int(request.GET['desde']),
                                                                 fecha_pago__year=int(request.GET['anio'])).exists()):
                        historial = HistorialPagoMes(
                            contrato_id=int(request.GET['id']),
                            fecha_pago=date(year=int(request.GET['anio']), month=int(request.GET['desde']),
                                            day=date.today().day),
                            cancelado=False
                        )
                        historial.save(request)
                        log('Agrega historial de pago %s' % (historial), request, 'add')
                    else:
                        historial = HistorialPagoMes.objects.get(status=True, contrato_id=int(request.GET['id']),
                                                                 fecha_pago__month=int(request.GET['desde']),
                                                                 fecha_pago__year=int(request.GET['anio']))
                    name_file = 'Informe_actividad_' + str(secinforme) + random.randint(1, 10000).__str__() + '.pdf'
                    folder = 'contratoepunemi/informe/'
                    try:
                        os.stat(os.path.join(SITE_STORAGE, 'media', folder))
                    except:
                        os.mkdir(os.path.join(SITE_STORAGE, 'media', folder))
                    secpos = \
                    SecuenciaMemoActividadPosgrado.objects.filter(tipo=PlantillaContratoDip.objects.last(), status=True,
                                                                  anioejercicio__anioejercicio=datetime.now().year)[0]
                    informe = InformeActividadJornada(mes=int(desde), contrato=contratopersona, secuenciageneral=secpos,
                                                      secuencia=secinforme, historialpago=historial)
                    # data = extraervalores(contratopersona.persona.id,request.GET['desde'],request.GET['anio'],True if 'h' in request.GET else False)
                    data['bitacora'] = BitacoraActividadDiaria.objects.filter(status=True,
                                                                              persona=contratopersona.persona,
                                                                              fecha__month=int(desde))
                    data['secuencia'] = '0' + str(informe) if informe.secuencia > 9 else '00' + str(informe)
                    data['dias'] = range(1, calendar.monthrange(int(request.GET['anio']), int(request.GET['desde']))[
                        1] + 1)
                    data['mes'] = request.GET['desde']
                    data['anio'] = request.GET['anio']
                    data['mes_nombre'] = MESES_CHOICES[int(desde) - 1][1]
                    pdf = conviert_html_to_pdf_save_informe('adm_contratodip/detalle_jornada_pdf.html',
                                                            {'pagesize': 'A4', 'data': data}, folder,
                                                            name_file)
                    if pdf[0]:
                        informe.archivo.name = os.path.join(folder, name_file)
                        informe.save(request)
                        mensaje = {"result": True, "mensaje": 'Informe generado correctamente'}
                    else:
                        transaction.set_rollback(True)
                        mensaje = {"result": False, "mensaje": 'Error al guardar pdf'}
                    return JsonResponse(mensaje)
                except Exception as ex:
                    print(ex)
                    transaction.set_rollback(True)
                    return JsonResponse({"result": False, "mensaje": "Error al procesar los datos."})

            elif action == 'inftecnico':
                try:
                    ids = request.GET['id']
                    desde = request.GET['desde']
                    data['persona'] = contratopersona = ContratoDip.objects.filter(pk=int(ids))[0]
                    fechas = TrabajadorDiaJornada.objects.values_list('fecha').filter(persona=contratopersona.persona,
                                                                                      anio=2022, mes=int(desde),
                                                                                      status=True).order_by('fecha')
                    # dias_no_laborable = DiasNoLaborable.objects.values_list('fecha').filter(fecha__in=fechas).exclude(periodo__isnull=False)
                    # data['dias'] = cumplehoras = TrabajadorDiaJornada.objects.filter(persona=contratopersona.persona, anio=2022, mes=int(desde),
                    #                                                                  status=True).exclude(fecha__in=dias_no_laborable).order_by('fecha')
                    if InformeTecnico.objects.values('id').filter(mes=int(desde), status=True,
                                                                  contrato=contratopersona).exists():
                        return JsonResponse({'result': False, 'mensaje': 'Ya existe el registro con el mes de %s' % (
                        MESES_CHOICES[int(desde) - 1][1])})
                    if not InformeActividadJornada.objects.values('id').filter(mes=int(desde), status=True,
                                                                               contrato=contratopersona).exists():
                        return JsonResponse(
                            {'result': False, 'mensaje': 'No ha generado el reporte de asistencia y actividades'})
                    if not (
                    HistorialPagoMes.objects.values('id').filter(status=True, contrato_id=int(request.GET['id']),
                                                                 fecha_pago__month=int(request.GET['desde']),
                                                                 fecha_pago__year=int(request.GET['anio'])).exists()):
                        historial = HistorialPagoMes(
                            contrato_id=int(request.GET['id']),
                            fecha_pago=date(year=int(request.GET['anio']), month=int(request.GET['desde']),
                                            day=date.today().day),
                            cancelado=False
                        )
                        historial.save(request)
                        log('Agrega historial de pago %s' % (historial), request, 'add')
                    else:
                        historial = HistorialPagoMes.objects.get(status=True, contrato_id=int(request.GET['id']),
                                                                 fecha_pago__month=int(request.GET['desde']),
                                                                 fecha_pago__year=int(request.GET['anio']))
                    data['info'] = InformeActividadJornada.objects.filter(mes=int(desde), status=True,
                                                                          contrato=contratopersona).last()
                    d = data['info'].archivo.name.split('\\')
                    data['nombrearchivo'] = d[-1].split('.')[0]
                    secinforme = contratopersona.secuencia_inftecnico()

                    name_file = 'Informe_Tecnico_' + str(secinforme) + random.randint(1, 10000).__str__() + '.pdf'
                    folder = 'contratoepunemi/inftec/'
                    try:
                        os.stat(os.path.join(SITE_STORAGE, 'media', folder))
                    except:
                        os.mkdir(os.path.join(SITE_STORAGE, 'media', folder))
                    secpos = \
                    SecuenciaMemoActividadPosgrado.objects.filter(tipo=PlantillaContratoDip.objects.last(), status=True,
                                                                  anioejercicio__anioejercicio=datetime.now().year)[0]
                    informe = InformeTecnico(mes=int(desde), contrato=contratopersona, secuenciageneral=secpos,
                                             secuencia=secinforme, historialpago=historial)
                    # data = extraervalores(contratopersona.persona.id,request.GET['desde'],request.GET['anio'],True if 'h' in request.GET else False)
                    data['fehca'] = date.today()
                    data['mes'] = MESES_CHOICES[datetime.today().month - 1]
                    data['actividades'] = BitacoraActividadDiaria.objects.values_list('actividades__descripcion',
                                                                                      flat=True).filter(status=True,
                                                                                                        fecha__month=int(
                                                                                                            desde),
                                                                                                        actividades__isnull=False,
                                                                                                        persona=persona).distinct()
                    data['secuencia'] = '0' + str(informe) if informe.secuencia > 9 else '00' + str(informe)
                    pdf = conviert_html_to_pdf_save_informe('adm_contratodip/informe_tecnico_pdf.html',
                                                            {'pagesize': 'A4', 'data': data}, folder,
                                                            name_file)
                    if pdf:
                        informe.archivo.name = os.path.join(folder, name_file)
                        informe.save(request)
                        mensaje = {"result": True, "mensaje": 'Informe generado correctamente'}
                    else:
                        mensaje = {"result": False, "mensaje": 'Error al guardar pdf'}
                    return JsonResponse(mensaje)
                except Exception as ex:
                    print(ex)
                    transaction.set_rollback(True)
                    return JsonResponse({"result": False, "mensaje": "Error al procesar los datos."})

            elif action == 'listcampoespecifico':
                try:
                    campoamplio = request.GET.get('campoamplio')
                    listcampoamplio = campoamplio
                    if len(campoamplio) > 1:
                        listcampoamplio = campoamplio.split(',')
                    querybase = SubAreaConocimientoTitulacion.objects.filter(status=True,
                                                                             areaconocimiento__in=listcampoamplio).order_by(
                        'codigo')
                    if 'q' in request.GET:
                        q = request.GET['q'].upper().strip()
                        if q != 'UNDEFINED':
                            querybase = querybase.filter((Q(nombre__icontains=q) | Q(codigo__icontains=q))).distinct()[
                                        :30]
                    data = {"result": "ok", "results": [
                        {"id": x.id, "idca": x.areaconocimiento.id, "name": "{} - {}".format(x.codigo, x.nombre)} for x
                        in querybase]}
                    return JsonResponse(data)
                except Exception as ex:
                    pass

            elif action == 'listcampodetallado':
                try:
                    campoespecifico = request.GET.get('campoespecifico')
                    listcampoespecifico = campoespecifico
                    if len(campoespecifico) > 1:
                        listcampoespecifico = campoespecifico.split(',')
                    querybase = SubAreaEspecificaConocimientoTitulacion.objects.filter(status=True,
                                                                                       areaconocimiento__in=listcampoespecifico).order_by(
                        'codigo')
                    if 'q' in request.GET:
                        q = request.GET['q'].upper().strip()
                        if q != 'UNDEFINED':
                            querybase = querybase.filter((Q(nombre__icontains=q) | Q(codigo__icontains=q))).distinct()[
                                        :30]
                    data = {"result": "ok",
                            "results": [{"id": x.id, "name": "{} - {}".format(x.codigo, x.nombre)} for x in querybase]}
                    return JsonResponse(data)
                except Exception as ex:
                    pass

            elif action == 'addarchivoevidencias':
                try:
                    if int(request.GET['tipo']) == 1:
                        arch = MemoActividadPosgrado.objects.get(status=True, pk=int(encrypt(request.GET['id'])))
                    elif int(request.GET['tipo']) == 2:
                        arch = InformeActividadJornada.objects.get(status=True, pk=int(encrypt(request.GET['id'])))
                    elif int(request.GET['tipo']) == 3:
                        arch = InformeTecnico.objects.get(status=True, pk=int(encrypt(request.GET['id'])))
                    else:
                        arch = ActaPago.objects.get(status=True, pk=int(encrypt(request.GET['id'])))
                    if arch.archivofirmado:
                        data['filtro'] = arch
                        data['id'] = encrypt(request.GET['id'])
                        data['tipo'] = request.GET['tipo']
                        data['form2'] = ArchivoInformesForm(initial={'archivo': arch.archivofirmado})
                    else:
                        form2 = ArchivoInformesForm()
                        data['tipo'] = request.GET['tipo']
                        data['id'] = encrypt(request.GET['id'])
                        data['form2'] = form2
                    template = get_template('adm_contratodip/modal/addarchivosinfo.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'traspasoconfirmado':
                try:
                    data['title'] = u'Confirmar Solicitud'
                    data['responsable'] = int(request.GET['responsable'])
                    data['solicitud'] = int(request.GET['solicitud'])
                    data['traspaso'] = ActivoFijo.objects.get(pk=int(request.GET['id']))
                    return render(request, "th_hojavida/traspasoconfirmado.html", data)
                except:
                    pass

            elif action == 'confmantenimiento':
                try:
                    data['title'] = u'Confirmar mantenimiento'
                    data['mantenimiento'] = SolicitudConfirmacionMantenimiento.objects.get(pk=int(request.GET['id']))
                    return render(request, "th_hojavida/modal/confmantenimiento.html", data)
                except:
                    pass

            elif action == 'rechmantenimiento':
                try:
                    data['title'] = u'Rechazar mantenimiento'
                    data['mantenimiento'] = SolicitudConfirmacionMantenimiento.objects.get(pk=int(request.GET['id']))
                    return render(request, "th_hojavida/modal/rechmantenimiento.html", data)
                except:
                    pass

            elif action == 'responsableconfirmatraspaso':
                try:
                    data['title'] = u'Confirmar Solicitud'
                    data['responsable'] = int(request.GET['responsable'])
                    data['solicitud'] = int(request.GET['solicitud'])
                    data['traspaso'] = ActivoFijo.objects.get(pk=int(request.GET['id']))
                    return render(request, "th_hojavida/responsableconfirmatraspaso.html", data)
                except:
                    pass

            elif action == 'traspasoeliminar':
                try:
                    data['title'] = u'Confirmar eliminación'
                    data['traspaso'] = int(request.GET['id'])
                    return render(request, "th_hojavida/traspasoeliminar.html", data)
                except:
                    pass

            elif action == 'responsableeliminatraspaso':
                try:
                    data['title'] = u'Confirmar eliminación'
                    data['traspaso'] = int(request.GET['id'])
                    return render(request, "th_hojavida/responsableeliminatraspaso.html", data)
                except:
                    pass

            elif action == 'addtransporte':
                try:
                    form = PersonaTransporteForm()
                    data['form'] = form
                    template = get_template('th_hojavida/modal/formtransporte.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'edittransporte':
                try:
                    data['transporte'] = transporte = PersonaTransporte.objects.get(id=int(encrypt(request.GET['id'])))
                    form = PersonaTransporteForm(initial=model_to_dict(transporte))
                    data['form'] = form
                    template = get_template('th_hojavida/modal/formtransporte.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'addalimentacion':
                try:
                    form = PersonaAlimentacionUniversidadForm()
                    data['form'] = form
                    template = get_template('th_hojavida/modal/formalimentacion.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'editalimentacion':
                try:
                    data['alimentacion'] = gastoalimentacion = PersonaAlimentacionUniversidad.objects.get(
                        id=int(encrypt(request.GET['id'])))
                    form = PersonaAlimentacionUniversidadForm(initial=model_to_dict(gastoalimentacion))
                    data['form'] = form
                    template = get_template('th_hojavida/modal/formalimentacion.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'addplantelefonico':
                try:
                    form = PersonaPlanTelefonicoForm()
                    data['form'] = form
                    template = get_template('th_hojavida/modal/formplantelefonico.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'editplantelefonico':
                try:
                    data['plantelefonico'] = plantelefonico = PersonaPlanTelefonico.objects.get(
                        id=int(encrypt(request.GET['id'])))
                    form = PersonaPlanTelefonicoForm(initial=model_to_dict(plantelefonico))
                    data['form'] = form
                    template = get_template('th_hojavida/modal/formplantelefonico.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'addcompraalimentos':
                try:
                    form = PersonaCompraAlimentosForm()
                    data['form'] = form
                    template = get_template('th_hojavida/modal/formcompraalimentos.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'editcompraalimentos':
                try:
                    data['compraalimento'] = compraalimento = PersonaCompraAlimentos.objects.get(
                        id=int(encrypt(request.GET['id'])))
                    form = PersonaCompraAlimentosForm(initial=model_to_dict(compraalimento))
                    data['form'] = form
                    template = get_template('th_hojavida/modal/formcompraalimentos.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'addgastoestudio':
                try:
                    form = PersonaGastoMensualForm()
                    data['form'] = form
                    template = get_template('th_hojavida/modal/formgastomensual.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'editgastoestudio':
                try:
                    data['gasto'] = gasto = PersonaGastoMensual.objects.get(
                        id=int(encrypt(request.GET['id'])))
                    form = PersonaGastoMensualForm(initial=model_to_dict(gasto))
                    data['form'] = form
                    template = get_template('th_hojavida/modal/formgastomensual.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'adddocumentofirmado':
                try:
                    data['filtro'] = filtro = PersonaAcciones.objects.get(pk=int(request.GET['id']))
                    data['formmodal'] = AccionPersonalDocumentoForm()
                    data['action'] = action
                    template = get_template("th_hojavida/modal/formdocumentofirmadobs4.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'historialfirmados':
                try:
                    data['filtro'] = filtro = PersonaAcciones.objects.get(pk=int(request.GET['id']))
                    data['action'] = action
                    template = get_template("th_hojavida/modal/historialdocumentofirmadobs4.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as e:
                    print(e)
                    pass

            elif action == 'addhistoriallaboralaspirante':
                try:
                    form2 = PersonaAportacionHistorialLaboralAspiranteForm()
                    data['form2'] = form2
                    data['accion'] = 'addhistoriallaboralaspirante'
                    template = get_template("th_hojavida/addhistoriallaboralaspirante.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'edithistoriallaboralaspirante':
                try:
                    form2 = PersonaAportacionHistorialLaboralAspiranteForm()
                    data['id'] = request.GET['id']
                    data['form2'] = form2
                    data['accion'] = 'addhistoriallaboralaspirante'
                    template = get_template("th_hojavida/addhistoriallaboralaspirante.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'historialfirmadosvacaciones':
                try:
                    data['filtro'] = filtro = AccionPersonal.objects.get(pk=int(request.GET['id']))
                    template = get_template("th_hojavida/modal/historialdocumentofirmadobs4.html")
                    data['action'] = action
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as e:
                    print(e)
                    pass

            elif action == 'importarActividades':
                try:
                    data['id_persona'] = request.GET['id']
                    data['action'] = 'importarActividades'
                    template = get_template("th_hojavida/modal/importarActividades.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

            # ACCIÓN DE PERSONAL
            elif action == 'firmadocvacaciones':
                try:
                    acc_personal = AccionPersonal.objects.get(id=request.GET['id'])
                    data['archivo'] = archivo = acc_personal.archivo.url
                    data['url_archivo'] = '{}{}'.format(dominio_sistema, archivo)
                    data['id_objeto'] = acc_personal.id
                    data['action_firma'] = 'firmadocvacaciones'
                    template = get_template("formfirmaelectronica.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'firmaraccionpersonal':
                try:
                    data['id'] = encrypt_id(request.GET['id'])
                    template = get_template("formfirmaelectronicamasiva.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')
                    return HttpResponseRedirect(request.path)

            elif action == 'accionpersonal':
                try:
                    data['title'] = u'Acción de personal'
                    url_vars = f'&action={action}'
                    acciones = persona.accionpersonal_set.filter(status=True).order_by('-id')
                    acciones2 = persona.personaacciones_set.filter(status=True).order_by('-id')
                    paging = MiPaginador(acciones, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['url_vars'] = url_vars
                    data['listado'] = page.object_list
                    data['listado2'] = acciones2
                    request.session['viewactivoth'] = ['informacionlaboral', action]
                    return render(request, "th_hojavida/informacionlaboral/accionpersonal.html", data)
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'adddocumentofirmadovacaciones':
                try:
                    data['filtro'] = filtro = AccionPersonal.objects.get(pk=int(request.GET['id']))
                    data['form'] = AccionPersonalDocumentoForm()
                    data['action'] = action
                    template = get_template("th_hojavida/modal/formdocumentofirmadobs4.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'editdocumentofirmadovacaciones':
                try:
                    data['filtro'] = filtro = AccionPersonal.objects.get(pk=int(request.GET['id']))
                    data['formmodal'] = AccionPersonalDocumentoForm()
                    data['action'] = action
                    template = get_template("th_hojavida/modal/formdocumentofirmadobs4.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            # DATOS PERSONALES

            elif action == 'editdatospersonales':
                try:
                    data['title'] = u'Datos personales'
                    persona = Persona.objects.get(id=persona.id)
                    perfil=persona.mi_perfil()
                    form = DatosPersonalesForm(initial={'nombres': persona.nombres,
                                                        'apellido1': persona.apellido1,
                                                        'apellido2': persona.apellido2,
                                                        'cedula': persona.cedula,
                                                        'pasaporte': persona.pasaporte,
                                                        'extension': persona.telefonoextension,
                                                        # 'agendacitas': persona.agendacitas,
                                                        'sexo': persona.sexo,
                                                        'lgtbi': persona.lgtbi,
                                                        'anioresidencia': persona.anioresidencia,
                                                        'nacimiento': persona.nacimiento,
                                                        'email': persona.email,
                                                        'estadocivil': persona.estado_civil(),
                                                        'libretamilitar': persona.libretamilitar,
                                                        'eszurdo': persona.eszurdo,
                                                        'estadogestacion': persona.estadogestacion,
                                                        'archivocedula':persona.documentos_personales().cedula if persona.documentos_personales() else '',
                                                        'papeleta':persona.archivo_papeleta() if persona.documentos_personales() else '',
                                                        'archivolibretamilitar':persona.archivo_libreta_militar() if persona.documentos_personales() else '',
                                                        'raza': perfil.raza,
                                                        'nacionalidadindigena': perfil.nacionalidadindigena,
                                                        'archivoraza': perfil.archivoraza
                                                        })
                    form.editar()
                    if esestudiante:
                        form.es_estudiante()
                    data['form'] = form
                    data['persona'] = persona
                    banderalibreta = 0
                    banderapapeleta = 0
                    banderacedula = 0
                    documentos = PersonaDocumentoPersonal.objects.filter(persona=persona)
                    if documentos:
                        if documentos[0].libretamilitar:
                            banderalibreta = 1
                        if documentos[0].papeleta:
                            banderapapeleta = 1
                        if documentos[0].cedula:
                            banderacedula = 1

                    data['banderacedula'] = banderacedula
                    data['banderalibreta'] = banderalibreta
                    data['banderapapeleta'] = banderapapeleta
                    data['switchery']=True
                    template = get_template('th_hojavida/informacionpersonal/modal/formdatospersonales.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'editdatosnacimiento':
                try:
                    data['title'] = u'Datos de nacimiento'
                    data['nacionalidad'] = persona.paisnacimiento.nacionalidad
                    form = DatosNacimientoForm(initial={'paisnacimiento': persona.paisnacimiento,
                                                        'provincianacimiento': persona.provincianacimiento,
                                                        'cantonnacimiento': persona.cantonnacimiento,
                                                        'parroquianacimiento': persona.parroquianacimiento,
                                                        'paisnacionalidad': persona.paisnacionalidad
                                                        })
                    form.editar(persona)
                    data['form'] = form
                    template = get_template('th_hojavida/informacionpersonal/modal/formnacimiento.html')
                    return JsonResponse({'result':True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'editdatosdomicilio':
                try:
                    data['title'] = u'Datos de domicilio'
                    form = DatosDomicilioForm(initial={'pais': persona.pais,
                                                       'provincia': persona.provincia,
                                                       'canton': persona.canton,
                                                       'ciudadela': persona.ciudadela,
                                                       'parroquia': persona.parroquia,
                                                       'direccion': persona.direccion,
                                                       'direccion2': persona.direccion2,
                                                       'sector': persona.sector,
                                                       'num_direccion': persona.num_direccion,
                                                       'referencia': persona.referencia,
                                                       'telefono': persona.telefono,
                                                       'telefono_conv': persona.telefono_conv,
                                                       'tipocelular': persona.tipocelular,
                                                       'archivoplanillaluz':persona.archivoplanillaluz,
                                                       'archivocroquis':persona.archivocroquis,
                                                       'serviciosbasico':persona.documentos_personales().serviciosbasico if persona.documentos_personales() else None,
                                                       'zona': persona.zona})
                    form.editar(persona)
                    data['form'] = form
                    template = get_template('th_hojavida/informacionpersonal/modal/formdomicilio.html')
                    return JsonResponse({'result':True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'datosmigrante':
                try:
                    data['title'] = u'Dirección de estudiantes en el extranjero'
                    if MigrantePersona.objects.filter(persona=persona).exists():
                        migrante = MigrantePersona.objects.get(persona=persona)
                        form = MigranteForm(initial={'paisresidencia': migrante.paisresidencia,
                                                     'anioresidencia': migrante.anioresidencia,
                                                     'mesresidencia': migrante.mesresidencia,
                                                     'archivo': migrante.archivo,
                                                     'fecharetorno': migrante.fecharetorno}
                                            )
                        tienearchivo = True if migrante.archivo else False
                    else:
                        form = MigranteForm()
                        tienearchivo = False

                    data['form'] = form
                    template = get_template('ajaxformmodal.html')
                    return JsonResponse({'result': True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            # DATOS FAMILIARES

            elif action == 'consultacedula':
                try:
                    cedula = request.GET['cedula'].strip().upper()
                    id=request.GET.get('id',0)
                    datospersona = Persona.objects.filter(Q(pasaporte=cedula) | Q(cedula=cedula) | Q(pasaporte=('VS'+cedula)) | Q(cedula=cedula[2:]), status=True).first()
                    if datospersona:
                        if datospersona == persona:
                            return JsonResponse({"result": "bad", 'mensaje': 'No puede agregar su propia cédula.'})
                        if persona.personadatosfamiliares_set.filter(identificacion=cedula, status=True).exclude(id=id).exists():
                            return JsonResponse({"result": "bad", 'mensaje':'Identificación ingresada ya se encuentra registrada en un familiar'})
                        perfil_i=datospersona.perfilinscripcion_set.filter(status=True).first()
                        editdiscapacidad=False
                        if len(datospersona.mis_perfilesusuarios()) == 1 and datospersona.tiene_usuario_externo():
                            editdiscapacidad = True
                        elif perfil_i and not perfil_i.estadoarchivodiscapacidad == 2:
                            editdiscapacidad = True
                        context={}
                        if perfil_i and perfil_i.tienediscapacidad:
                            context={'tipodiscapacidad':perfil_i.tipodiscapacidad.id if perfil_i.tipodiscapacidad else False,
                                    'tienediscapacidad':perfil_i.tienediscapacidad,
                                    'porcientodiscapacidad':perfil_i.porcientodiscapacidad,
                                    'carnetdiscapacidad':perfil_i.carnetdiscapacidad,
                                    'institucionvalida':perfil_i.institucionvalida.id if perfil_i.institucionvalida else False,
                                    'ceduladiscapacidad':perfil_i.archivo.url if perfil_i.archivo else False ,
                                    'archivoautorizado':perfil_i.archivovaloracion.url if perfil_i.archivovaloracion else False,
                                    }
                        return JsonResponse({"result": "ok",
                                             "apellido1": datospersona.apellido1,
                                             "apellido2": datospersona.apellido2,
                                             "nacimiento": datospersona.nacimiento.strftime('%Y-%m-%d'),
                                             "nombres": datospersona.nombres,
                                             "telefono": datospersona.telefono,
                                             "telefono_conv": datospersona.telefono_conv,
                                             "sexo": datospersona.sexo.id if datospersona.sexo else '' ,
                                             "puedeeditar": editdiscapacidad,
                                             "perfil_i":context})
                    else:
                        return JsonResponse({"result": "no"})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": str(ex)})

            elif action == 'datosfamiliares':
                try:
                    data['title'] = u'Datos familiares'
                    url_vars=f'&action={action}'
                    familiares = persona.familiares().order_by('-id')
                    paging = MiPaginador(familiares, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['url_vars'] = url_vars
                    data['listado'] = page.object_list
                    request.session['viewactivoth'] = ['informacionpersonal',action]
                    return render(request, "th_hojavida/informacionpersonal/datosfamiliares.html", data)
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'addfamiliar':
                try:
                    form = FamiliarForm()
                    visible_fields = form.visible_fields()
                    total_fields = len(visible_fields)
                    lista=[(1,'Informació Básica', visible_fields[:19]),
                             (2,'Información Laboral', visible_fields[19:27]),
                             (3,'Discapacidad',visible_fields[27:total_fields])
                            ]
                    data['form'] = lista
                    data['switchery']=True
                    template = get_template('th_hojavida/informacionpersonal/modal/formfamiliar.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'editfamiliar':
                try:
                    data['title'] = u'Editar familiar'
                    data['id'] = id = encrypt_id(request.GET['id'])
                    data['familiar'] = familiar = PersonaDatosFamiliares.objects.get(pk=id)
                    apellido1, apellido2, nombres, editdiscapacidad, sexo = '', '', familiar.nombre, False, None
                    if familiar.personafamiliar:
                        apellido1 = familiar.personafamiliar.apellido1
                        apellido2 = familiar.personafamiliar.apellido2
                        nombres = familiar.personafamiliar.nombres
                        sexo = familiar.personafamiliar.sexo
                        perfil_f=familiar.personafamiliar.perfilinscripcion_set.filter(status=True).first()
                        estado = True if perfil_f and perfil_f.estadoarchivodiscapacidad == 2 else False
                        if len(familiar.personafamiliar.mis_perfilesusuarios()) == 1 and familiar.personafamiliar.tiene_usuario_externo():
                            editdiscapacidad = True
                        elif not estado:
                            editdiscapacidad = True
                    else:
                        editdiscapacidad = True
                    banderacedula = 0
                    if familiar.cedulaidentidad:
                        banderacedula = 1
                    data['banderacedula'] = banderacedula
                    data['edit_d'] = editdiscapacidad
                    form = FamiliarForm(initial={'identificacion': familiar.identificacion,
                                                 'parentesco': familiar.parentesco,
                                                 'nombre': nombres,
                                                 'apellido1': apellido1,
                                                 'apellido2': apellido2,
                                                 'sexo': sexo,
                                                 'nacimiento': familiar.nacimiento,
                                                 'fallecido': familiar.fallecido,
                                                 'tienediscapacidad': familiar.tienediscapacidad,
                                                 'telefono': familiar.telefono,
                                                 'niveltitulacion': familiar.niveltitulacion,
                                                 'ingresomensual': familiar.ingresomensual,
                                                 'formatrabajo': familiar.formatrabajo,
                                                 'telefono_conv': familiar.telefono_conv,
                                                 'trabajo': familiar.trabajo,
                                                 'convive': familiar.convive,
                                                 'sustentohogar': familiar.sustentohogar,
                                                 'essustituto': familiar.essustituto,
                                                 'autorizadoministerio': familiar.autorizadoministerio,
                                                 'tipodiscapacidad': familiar.tipodiscapacidad,
                                                 'porcientodiscapacidad': familiar.porcientodiscapacidad,
                                                 'carnetdiscapacidad': familiar.carnetdiscapacidad,
                                                 'institucionvalida': familiar.institucionvalida,
                                                 'tipoinstitucionlaboral': familiar.tipoinstitucionlaboral,
                                                 'tienenegocio': familiar.tienenegocio,
                                                 'esservidorpublico': familiar.esservidorpublico,
                                                 'bajocustodia': familiar.bajocustodia,
                                                 'tienenegocio': familiar.tienenegocio,
                                                 'cedulaidentidad': familiar.cedulaidentidad,
                                                 'ceduladiscapacidad': familiar.ceduladiscapacidad,
                                                 'archivoautorizado': familiar.archivoautorizado,
                                                 'cartaconsentimiento': familiar.cartaconsentimiento,
                                                 'archivocustodia': familiar.archivocustodia,
                                                 'centrocuidado': familiar.centrocuidado,
                                                 'centrocuidadodesc': familiar.centrocuidadodesc,
                                                 'negocio': familiar.negocio, })
                    visible_fields = form.visible_fields()
                    total_fields = len(visible_fields)
                    lista = [(1, 'Informació Básica', visible_fields[:19]),
                             (2, 'Información Laboral', visible_fields[19:27]),
                             (3, 'Discapacidad', visible_fields[27:total_fields])
                             ]
                    form.edit()
                    data['form'] = lista
                    data['switchery'] = True
                    template = get_template('th_hojavida/informacionpersonal/modal/formfamiliar.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'subiractacargafamiliar':
                try:
                    data['id_persona'] = request.GET['id']
                    template = get_template("th_hojavida/modal/formactafirmada.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'firmaactaproy':
                try:
                    archivoacta = Archivo.objects.get(id=request.GET['id'])
                    data['archivo'] = archivo = archivoacta.archivo.url
                    data['url_archivo'] = '{}{}'.format(dominio_sistema, archivo)
                    data['id_objeto'] = archivoacta.id
                    data['action_firma'] = 'firmaactaproy'
                    template = get_template("formfirmaelectronica.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            # DATOS FINANCIEROS

            elif action == 'finanzas':
                try:
                    data['title'] = u'Cuenta bancaria'
                    data['cuentas'] = persona.cuentasbancarias()
                    request.session['viewactivoth'] = ['informacionpersonal', action]
                    return render(request, "th_hojavida/informacionpersonal/finanzas.html", data)
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'contraloria':
                try:
                    data['title'] = u'Declaración patrimonial (Contraloría)'
                    data['declaraciones'] = persona.declaracionesbienes().order_by('-fecha')
                    request.session['viewactivoth'] = ['informacionpersonal', action]
                    return render(request, "th_hojavida/informacionpersonal/contraloria.html", data)
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'addcuentabancaria':
                try:
                    form = CuentaBancariaPersonaForm()
                    data['form'] = form
                    data['switchery'] = True
                    template = get_template('th_hojavida/informacionmedica/modal/formbasico.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'editcuentabancaria':
                try:
                    data['title'] = u'Editar cuenta bancaria'
                    data['cuentabancaria'] = cuentabancaria = CuentaBancariaPersona.objects.get(
                        pk=int(request.GET['id']))
                    data['id'] = request.GET['id']
                    data['form'] = CuentaBancariaPersonaForm(initial={'numero': cuentabancaria.numero,
                                                                      'banco': cuentabancaria.banco,
                                                                      'tipocuentabanco': cuentabancaria.tipocuentabanco})
                    template = get_template('th_hojavida/informacionmedica/modal/formbasico.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'adddeclaracion':
                try:
                    form = DeclaracionBienForm()
                    ids_d = DistributivoPersona.objects.filter(status=True, persona=persona).values_list('denominacionpuesto_id', flat=True)
                    cargos = DenominacionPuesto.objects.filter(id__in=ids_d)
                    data['form'] = form
                    form.fields['cargo'].queryset = cargos
                    if cargos:
                        form.fields['cargo'].initial = cargos[0]
                    data['switchery'] = True
                    template = get_template('th_hojavida/informacionpersonal/modal/adddeclaracion.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'editdeclaracion':
                try:
                    data['id'] = id = encrypt_id(request.GET['id'])
                    declaracion = DeclaracionBienes.objects.get(pk=id)
                    form = DeclaracionBienForm(initial={
                        'tipodeclaracion': declaracion.tipodeclaracion,
                        'fecha': declaracion.fecha,
                        'codigobarra': declaracion.codigobarra,
                        'fechaperiodoinicio': declaracion.fechaperiodoinicio,
                        'archivo': declaracion.archivo,
                        # 'cargo' : declaracion.denominacionpuesto,
                    })

                    ids_d = DistributivoPersona.objects.filter(status=True, persona=persona).values_list('denominacionpuesto_id', flat=True)

                    if declaracion.denominacionpuesto:
                        if declaracion.denominacionpuesto.id not in ids_d:
                            ids_cargos_his = DistributivoPersonaHistorial.objects.filter(status=True, persona=persona).values_list('denominacionpuesto_id', flat=True).distinct()
                            if declaracion.denominacionpuesto.id  in ids_cargos_his:
                                form.fields['cargosvigentes'].initial = False
                                ids_d = ids_cargos_his
                            else:
                                ids_d = []
                                form.fields['cargosvigentes'].initial = False
                                form.fields['cargo'].initial = declaracion.denominacionpuesto.id

                    cargos = DenominacionPuesto.objects.filter(id__in=ids_d)

                    if cargos:
                        form.fields['cargo'].queryset = cargos
                    if declaracion.denominacionpuesto:
                        form.fields['cargo'].initial = declaracion.denominacionpuesto.id

                    if declaracion.departamento and declaracion.tipodeclaracion == 1:
                        form.fields['departamento'].initial = declaracion.departamento.id

                    data['form'] = form
                    form.fields['archivo'].required = False
                    # form.fields['cargo'].queryset = cargos
                    # form.fields['cargo'].initial = declaracion.denominacionpuesto.id
                    data['switchery'] = True
                    template = get_template('th_hojavida/informacionpersonal/modal/adddeclaracion.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'referencias':
                try:
                    data['title'] = u'Referencias personales'
                    data['listado'] = persona.referenciapersona_set.filter(status=True)
                    request.session['viewactivoth'] = ['informacionpersonal', action]
                    return render(request, "th_hojavida/informacionpersonal/referencias.html", data)
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'addreferencia':
                try:
                    form = ReferenciaPersonaForm()
                    data['form'] = form
                    data['switchery'] = True
                    template = get_template('th_hojavida/informacionmedica/modal/formbasico.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'editreferencia':
                try:
                    data['title'] = u'Editar referencia personal'
                    data['referencia'] = referencia = ReferenciaPersona.objects.get(pk=request.GET['id'])
                    form = ReferenciaPersonaForm(initial={'nombres': referencia.nombres,
                                                          'apellidos': referencia.apellidos,
                                                          'email': referencia.email,
                                                          'telefono': referencia.telefono,
                                                          'institucion': referencia.institucion,
                                                          'cargo': referencia.cargo,
                                                          'relacion': referencia.relacion})
                    data['form'] = form
                    data['id'] = request.GET['id']
                    template = get_template('th_hojavida/informacionmedica/modal/formbasico.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            # FORMACIÓN ACADEMICA

            elif action == 'academica':
                try:
                    data['title'] = u'Formación académica'
                    url_vars=f'&action={action}'
                    titulaciones = persona.mis_titulaciones().order_by('-titulo__niveltitulacion')
                    paging = MiPaginador(titulaciones, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['url_vars'] = url_vars
                    data['listado'] = page.object_list
                    request.session['viewactivoth'] = ['informacionacademica',action]
                    return render(request, "th_hojavida/informacionacademica/formacionacademica.html", data)
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'addtitulacion':
                try:
                    data['title'] = u'Adicionar titulación'
                    form = TitulacionPersonaForm()
                    form.adicionar()
                    visible_fields = form.visible_fields()
                    total_fields = len(visible_fields)
                    lista = [(1, 'Información de título', visible_fields[:14]),
                             (2, 'Ubicación de institución', visible_fields[14:18]),
                             (3, 'Beca', visible_fields[18:total_fields])
                             ]
                    data['form'] = lista
                    data['switchery'] = True
                    template = get_template('th_hojavida/informacionacademica/modal/formtitulacion.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'edittitulacion':
                try:
                    data['titulacion'] = titulacion = Titulacion.objects.get(pk=encrypt_id(request.GET['id']))
                    # campotitulo, campoamplio, campoespecifico, campodetallado = None, None, None, None
                    # if CamposTitulosPostulacion.objects.filter(status=True, titulo=titulacion.titulo).exists():
                    #     campotitulo = CamposTitulosPostulacion.objects.filter(status=True,
                    #                                                           titulo=titulacion.titulo).first()
                    #     campoamplio = AreaConocimientoTitulacion.objects.filter(status=True,
                    #                                                             id__in=campotitulo.campoamplio.all().values_list(
                    #                                                                 'id', flat=True))
                    #     campoespecifico = SubAreaConocimientoTitulacion.objects.filter(status=True,
                    #                                                                    id__in=campotitulo.campoespecifico.all().values_list(
                    #                                                                        'id', flat=True))
                    #     campodetallado = SubAreaEspecificaConocimientoTitulacion.objects.filter(status=True,
                    #                                                                             id__in=campotitulo.campodetallado.all().values_list(
                    #                                                                                 'id', flat=True))
                    form = TitulacionPersonaForm(initial={'titulo': titulacion.titulo,
                                                          'areatitulo': titulacion.areatitulo,
                                                          'fechainicio': titulacion.fechainicio,
                                                          'educacionsuperior': titulacion.educacionsuperior,
                                                          'institucion': titulacion.institucion,
                                                          'colegio': titulacion.colegio,
                                                          'cursando': titulacion.cursando,
                                                          'fechaobtencion': titulacion.fechaobtencion if not titulacion.cursando else datetime.now().date(),
                                                          'fechaegresado': titulacion.fechaegresado if not titulacion.cursando else datetime.now().date(),
                                                          'registro': titulacion.registro,
                                                          # 'areaconocimiento': titulacion.areaconocimiento,
                                                          # 'subareaconocimiento': titulacion.subareaconocimiento,
                                                          # 'subareaespecificaconocimiento': titulacion.subareaespecificaconocimiento,
                                                          'pais': titulacion.pais,
                                                          'provincia': titulacion.provincia,
                                                          'canton': titulacion.canton,
                                                          'parroquia': titulacion.parroquia,
                                                          # 'campoamplio': campoamplio,
                                                          # 'campoespecifico': campoespecifico,
                                                          # 'campodetallado': campodetallado,
                                                          'anios': titulacion.anios,
                                                          'semestres': titulacion.semestres,
                                                          'aplicobeca': titulacion.aplicobeca,
                                                          'tipobeca': titulacion.tipobeca,
                                                          'archivo': titulacion.archivo,
                                                          'registroarchivo': titulacion.registroarchivo,
                                                          'financiamientobeca': titulacion.financiamientobeca,
                                                          'valorbeca': titulacion.valorbeca})
                    form.editar(titulacion)
                    visible_fields = form.visible_fields()
                    total_fields = len(visible_fields)
                    lista = [(1, 'Información de título', visible_fields[:14]),
                             (2, 'Ubicación de institución', visible_fields[14:18]),
                             (3, 'Beca', visible_fields[18:total_fields])
                             ]
                    data['form'] = lista
                    data['id'] = titulacion.id
                    data['switchery'] = True
                    template = get_template('th_hojavida/informacionacademica/modal/formtitulacion.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'buscartitulo':
                try:
                    q = request.GET['q'].upper().strip()
                    s = q.split(" ")
                    filtro = Q(status=True, nombre__icontains=q)
                    titulos = Titulo.objects.filter(filtro).distinct()[:15]
                    resp = [{'id': t.pk, 'text': f"{t.nombre}"} for t in titulos]
                    return JsonResponse(resp, safe=False)
                except Exception as ex:
                    pass

            elif action == 'detalletitulo':
                try:
                    data['titulacion'] = titulacion = Titulacion.objects.get(pk=encrypt_id(request.GET['id']))
                    dettitu = titulacion.detalletitulacionbachiller_set.filter(status=True)
                    data['detalletitulacionbachiller'] = dettitu.last()
                    if titulacion.usuario_creacion:
                        data['personacreacion'] = Persona.objects.get(
                            usuario=titulacion.usuario_creacion) if titulacion.usuario_creacion.id > 1 else ""
                    template = get_template("th_hojavida/detalletitulo.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": True, 'data': json_content})
                except Exception as ex:
                    return JsonResponse({"result": False, "mensaje": u"Error al obtener los datos."})

            elif action == 'adddetalletitulobachiller':
                try:
                    data['titulacion'] = titulacion = Titulacion.objects.get(pk=encrypt_id(request.GET['id']))
                    data['id']=titulacion.id
                    dettitu = titulacion.detalletitulacionbachiller_set.filter(status=True)
                    archivobachiller = Archivo.objects.filter(Q(inscripcion__persona=persona)|Q(persona=persona),
                                                              tipo_id=16,
                                                              status=True)
                    archivoreconocimiento = Archivo.objects.filter(Q(inscripcion__persona=persona)|Q(persona=persona),
                                                                   tipo_id=18,
                                                                   status=True)
                    # pasar archivos de detalletitulacion a Archivos y viceversa
                    form = None
                    if dettitu or archivobachiller or archivoreconocimiento:
                        detalle = dettitu.last() if dettitu else None

                        if not detalle and archivobachiller:
                            detalle = DetalleTitulacionBachiller(
                                titulacion=titulacion,
                                actagrado=archivobachiller.last().archivo
                            )
                            detalle.save(request)

                        elif not detalle and archivoreconocimiento:
                            detalle = DetalleTitulacionBachiller(
                                titulacion=titulacion,
                                reconocimientoacademico=archivoreconocimiento.last().archivo
                            )
                            detalle.save(request)

                        elif detalle.actagrado and not archivobachiller:
                            inscri = perfilprincipal.inscripcion if perfilprincipal.es_estudiante() else None
                            profe = perfilprincipal.inscripcion if perfilprincipal.es_profesor() else None
                            archivobachiller = Archivo(
                                tipo_id=16,
                                nombre=f'ACTA DE GRADO DE BACHILLER DE LA PERSONA: {persona}',
                                fecha=datetime.now().date(),
                                archivo=dettitu.last().actagrado,
                                aprobado=True,
                                inscripcion=inscri,
                                profesor=profe,
                                persona=persona,
                                sga=True
                            )
                            archivobachiller.save(request)

                        elif detalle.reconocimientoacademico and not archivoreconocimiento:
                            inscri = perfilprincipal.inscripcion if perfilprincipal.es_estudiante() else None
                            profe = perfilprincipal.inscripcion if perfilprincipal.es_profesor() else None
                            archivoreconocimiento = Archivo(
                                tipo_id=18,
                                nombre=f'RECONOCIMIENTO ACADÉMICO DE LA PERSONA: {persona}',
                                fecha=datetime.now().date(),
                                archivo=dettitu.last().reconocimientoacademico,
                                aprobado=True,
                                inscripcion=inscri,
                                profesor=profe,
                                persona=persona,
                                sga=True
                            )
                            archivoreconocimiento.save(request)

                        form = DetalleTitulacionBachillerForm(
                            initial={'titulacion': titulacion.id if titulacion else "",
                                     'calificacion': detalle.calificacion,
                                     'archivo': detalle.actagrado,
                                     'anioinicioperiodograduacion': detalle.anioinicioperiodograduacion,
                                     'aniofinperiodograduacion': detalle.aniofinperiodograduacion,
                                     'reconocimientoacademico': detalle.reconocimientoacademico})
                    else:
                        form = DetalleTitulacionBachillerForm(
                            initial={'titulacion': titulacion.id if titulacion else ""})
                    # form.deshabilitar(titulacion)

                    data['form'] = form
                    template = get_template('th_hojavida/informacionacademica/modal/formtitulacionbachiller.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'articulos':
                try:
                    data['title'] = u'Artículos'
                    url_vars = f'&action={action}'

                    # Consulta los artículos
                    articulos = ArticuloInvestigacion.objects.select_related().filter((Q(
                        participantesarticulos__profesor__persona=persona) | Q(
                        participantesarticulos__administrativo__persona=persona) | Q(
                        participantesarticulos__inscripcion__persona=persona)),
                        status=True, aprobado=True,
                        participantesarticulos__status=True).order_by('-fechapublicacion')

                    paging = MiPaginador(articulos, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['url_vars'] = url_vars
                    data['listado'] = page.object_list
                    request.session['viewactivoth'] = ['informacionacademica', action]
                    return render(request, "th_hojavida/informacionacademica/articulos.html", data)
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'evidenciasarticulo':
                try:
                    data['articulo'] = articulos = ArticuloInvestigacion.objects.get(pk=int(encrypt(request.GET['id'])))
                    data['evidencias'] = Evidencia.objects.filter(status=True, matrizevidencia_id=3)

                    template = get_template("th_hojavida/detallearticulo.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    msg = ex.__str__()
                    return JsonResponse({"result": "bad", "message": u"Error al obtener los datos. [%s]" % msg})

            elif action == 'ponencias':
                try:
                    data['title'] = u'Ponencias'
                    url_vars = f'&action={action}'

                    # Consulta las ponencias
                    ponencias = PonenciasInvestigacion.objects.select_related().filter((Q(
                        participanteponencias__profesor__persona=persona) | Q(
                        participanteponencias__administrativo__persona=persona) | Q(
                        participanteponencias__inscripcion__persona=persona)),
                        status=True, participanteponencias__status=True)

                    paging = MiPaginador(ponencias, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['url_vars'] = url_vars
                    data['listado'] = page.object_list
                    request.session['viewactivoth'] = ['informacionacademica', action]
                    return render(request, "th_hojavida/informacionacademica/ponencias.html", data)
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'evidenciasponencia':
                try:
                    data['ponencias'] = ponencias = PonenciasInvestigacion.objects.get(pk=int(encrypt(request.GET['id'])))
                    data['evidencias'] = Evidencia.objects.filter(status=True, matrizevidencia_id=4)

                    template = get_template("th_hojavida/detalleponencia.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    msg = ex.__str__()
                    return JsonResponse({"result": "bad", "message": u"Error al obtener los datos. [%s]" % msg})

            elif action == 'libros':
                try:
                    data['title'] = u'Libros'
                    url_vars = f'&action={action}'

                    # Consulta los libros
                    libros = LibroInvestigacion.objects.select_related().filter((Q(
                        participantelibros__profesor__persona=persona) | Q(participantelibros__profesor__persona=persona)),
                        status=True, participantelibros__status=True)

                    paging = MiPaginador(libros, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['url_vars'] = url_vars
                    data['listado'] = page.object_list
                    request.session['viewactivoth'] = ['informacionacademica', action]
                    return render(request, "th_hojavida/informacionacademica/libros.html", data)
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'evidenciaslibro':
                try:
                    data['libros'] = libros = LibroInvestigacion.objects.get(pk=int(encrypt(request.GET['id'])))
                    data['evidencias'] = Evidencia.objects.filter(status=True, matrizevidencia_id=6)

                    template = get_template("th_hojavida/detalledetallelibro.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    msg = ex.__str__()
                    return JsonResponse({"result": "bad", "message": u"Error al obtener los datos. [%s]" % msg})

            elif action == 'capitulos':
                try:
                    data['title'] = u'Capítulos de libros'
                    url_vars = f'&action={action}'

                    # Consulta los capítulos
                    capitulos = CapituloLibroInvestigacion.objects.select_related().filter((Q(
                        participantecapitulolibros__profesor__persona=persona) | Q(
                        participantecapitulolibros__profesor__persona=persona)),
                        status=True, participantecapitulolibros__status=True)

                    paging = MiPaginador(capitulos, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['url_vars'] = url_vars
                    data['listado'] = page.object_list
                    request.session['viewactivoth'] = ['informacionacademica', action]
                    return render(request, "th_hojavida/informacionacademica/capitulos.html", data)
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'evidenciascapitulo':
                try:
                    data['capitulos'] = capitulos = CapituloLibroInvestigacion.objects.get(pk=int(encrypt(request.GET['id'])))
                    data['evidencias'] = Evidencia.objects.filter(status=True, matrizevidencia_id=7)

                    template = get_template("th_hojavida/detallecapitulo.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    msg = ex.__str__()
                    return JsonResponse({"result": "bad", "message": u"Error al obtener los datos. [%s]" % msg})

            elif action == 'redesacademicas':
                try:
                    data['title'] = u'Redes Académicas'
                    url_vars = f'&action={action}'

                    # Consulta las redes académicas
                    redesacademicas = RedPersona.objects.filter(persona=persona, status=True)

                    paging = MiPaginador(redesacademicas, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['url_vars'] = url_vars
                    data['listado'] = page.object_list
                    request.session['viewactivoth'] = ['informacionacademica', action]
                    return render(request, "th_hojavida/informacionacademica/redesacademicas.html", data)
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'addredacademica':
                try:
                    form = RedAcademicaForm()
                    data['form'] = form
                    template = get_template('th_hojavida/informacionacademica/modal/formredacademica.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'editredacademica':
                try:
                    redacademica = RedPersona.objects.get(pk=int(encrypt(request.GET['id'])))
                    data['id'] = redacademica.id
                    form = RedAcademicaForm(initial=model_to_dict(redacademica))
                    data['form'] = form
                    template = get_template('th_hojavida/informacionacademica/modal/formredacademica.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'revisionesarticulos':
                try:
                    data['title'] = u'Par revisor artículos'
                    url_vars = f'&action={action}'

                    # Consulta las revisiones de produccion científica
                    revisionesproduccion = ParRevisorProduccionCientifica.objects.filter(persona=persona, status=True).order_by('-fecharevision')

                    paging = MiPaginador(revisionesproduccion, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['url_vars'] = url_vars
                    data['listado'] = page.object_list
                    request.session['viewactivoth'] = ['informacionacademica', action]
                    return render(request, "th_hojavida/informacionacademica/revisionesarticulos.html", data)
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'addparrevisor':
                try:
                    form = ParRevisorArticuloForm()
                    data['form'] = form
                    template = get_template('th_hojavida/informacionacademica/modal/formparrevisor.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'editparrevisor':
                try:
                    parrevisor = ParRevisorProduccionCientifica.objects.get(pk=int(encrypt(request.GET['id'])))
                    data['id'] = parrevisor.id
                    form = ParRevisorArticuloForm(initial=model_to_dict(parrevisor))
                    data['form'] = form
                    template = get_template('th_hojavida/informacionacademica/modal/formparrevisor.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            # PROYECTOS

            elif action == 'proyectos':
                try:
                    data['title'] = u'Proyectos'
                    url_vars, tipo=f'&action={action}', request.GET.get('tipo','')
                    if not tipo:
                        proyectos = ParticipantesMatrices.objects.filter(matrizevidencia_id=2, status=True, proyecto__status=True, inscripcion__persona=persona)
                    elif tipo == 'e_investigacion':
                        proyectos = ProyectoInvestigacionExterno.objects.filter(persona=persona, status=True)
                    elif tipo == 'p_vinculacion':
                        ids_p, ids_a = [], []
                        if perfilprincipal.es_profesor():
                            if persona.profesor_set.filter(status=True).exists():
                                profesor = persona.profesor()
                                ids_p = profesor.participantesmatrices_set.filter(status=True, proyecto__status=True).values_list('id',flat=True)
                        admininstrativo= persona.administrativo_set.filter(status=True).first()
                        if admininstrativo:
                            ids_a = admininstrativo.participantesmatrices_set.filter(status=True, proyecto__status=True).values_list('id',flat=True)
                        ids=list(ids_p)+list(ids_a)
                        proyectos=ParticipantesMatrices.objects.filter(id__in=ids)

                    if tipo:
                        url_vars += '&tipo=' + tipo

                    paging = MiPaginador(proyectos, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['url_vars'] = url_vars
                    data['tipo'] = tipo
                    data['listado'] = page.object_list
                    request.session['viewactivoth'] = ['informacionacademica',action]
                    return render(request, "th_hojavida/informacionacademica/proyectos.html", data)
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'addproyectoexterno':
                try:
                    form = ProyectoInvestigacionExternoForm()
                    data['form'] = form
                    template = get_template('th_hojavida/informacionacademica/modal/formproyectoexterno.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'editproyectoexterno':
                try:
                    proyectoexterno = ProyectoInvestigacionExterno.objects.get(pk=int(encrypt(request.GET['id'])))
                    data['id'] = proyectoexterno.id
                    form = ProyectoInvestigacionExternoForm(initial=model_to_dict(proyectoexterno))
                    data['form'] = form
                    template = get_template('th_hojavida/informacionacademica/modal/formproyectoexterno.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            # TUTORIAS

            elif action == 'tutorias':
                try:
                    data['title'] = u'Tutorías'
                    data['subtitle'] = u'Listado de certificaciones de tutorías'
                    url_vars, tipo = f'&action={action}', request.GET.get('tipo','')
                    if not tipo:
                        listado=CertificadoTutoriaHV.objects.filter(status=True, persona=persona).order_by('id')
                    elif tipo == 'titulacionposgrado':
                        data['title'] = u'Tutorías de tesis posgrado'
                        data['subtitle'] = u'Listado de tutorías de tesis posgrado a cargo'
                        url_vars+=f'&tipo={tipo}'
                        profesor = persona.profesor()
                        listado=[]
                        if perfilprincipal.es_profesor() and profesor:
                            listado=TemaTitulacionPosgradoMatricula.objects.filter(status=True, tutor=profesor).order_by('pk')

                    paging = MiPaginador(listado, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['url_vars'] = url_vars
                    data['tipo'] = tipo
                    data['listado'] = page.object_list
                    request.session['viewactivoth'] = ['informacionacademica', action]
                    return render(request, "th_hojavida/informacionacademica/tutorias.html", data)
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'detallecertificaciontutoria':
                try:
                    data['certificacion'] = certificacion = CertificadoTutoriaHV.objects.get(pk=int(request.GET['id']))
                    if certificacion.usuario_creacion:
                        data['personacreacion'] = Persona.objects.get(usuario=certificacion.usuario_creacion) if certificacion.usuario_creacion.id > 1 else ""
                    template = get_template("th_hojavida/detallecertificaciontutoria.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, "mensaje": u"Error al obtener los datos."})

            elif action == 'addcertificadotutoria':
                try:
                    data['title'] = u'Adicionar Certificado de Tutoria'
                    data['persona'] = persona
                    form = CertificadoTutoriaForm()
                    data['form'] = form
                    template = get_template('ajaxformmodal.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'editcertificadotutoria':
                try:
                    data['title'] = u'Editar Certificado de Tutoria'
                    data['certificado'] = certificado = CertificadoTutoriaHV.objects.get(pk=encrypt_id(request.GET['id']))
                    data['id'] = certificado.id
                    data['form'] = CertificadoTutoriaForm(initial={'institucion': certificado.institucion,
                                                                   'nombreproyecto': certificado.nombreproyecto,
                                                                   'descripcion': certificado.descripcion,
                                                                   'calificacion': certificado.calificacion,
                                                                   'fechainicio': certificado.fechainicio,
                                                                   'archivo': certificado.archivo,
                                                                   'fechafin': certificado.fechafin}

                                                          )
                    template = get_template('ajaxformmodal.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            # DATOS MÉDICOS

            elif action == 'embarazo':
                try:
                    data['title'] = u'Gestación'
                    embarazos = persona.personadetallematernidad_set.filter(status=True).order_by('-pk')
                    data['listado'] = embarazos
                    request.session['viewactivoth'] = ['informacionmedica',action]
                    return render(request, "th_hojavida/informacionmedica/embarazo.html", data)
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'addembarazo':
                try:
                    form = PersonaDetalleMaternidadForm()
                    data['form'] = form
                    data['switchery'] = True
                    template = get_template('th_hojavida/informacionmedica/modal/formembarazo.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'editembarazo':
                try:
                    form = PersonaDetalleMaternidadForm()
                    data['id'] = id = request.GET['id']
                    detalle = PersonaDetalleMaternidad.objects.get(id=id)
                    form = PersonaDetalleMaternidadForm(initial=model_to_dict(detalle))
                    data['form'] = form
                    data['switchery'] = True
                    template = get_template('th_hojavida/informacionmedica/modal/formembarazo.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'datosmedicos':
                try:
                    data['title'] = u'Datos médicos'
                    data['subtitle'] = u'Información cargada'
                    data['url_vars'] = f'&action={action}'
                    data['persona'] = persona = Persona.objects.get(id=persona.id)
                    data['perfil'] = persona.mi_perfil()
                    data['datosextension'] =datosextension = persona.datos_extension()
                    data['examenfisico'] =examenfisico = persona.datos_examen_fisico()
                    data['vacunascovid'] = VacunaCovid.objects.filter(persona=persona, status=True).order_by('-id')
                    request.session['viewactivoth'] = ['informacionmedica',action]
                    return render(request, "th_hojavida/informacionmedica/datosmedicos.html", data)
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'editdatosmedicos':
                try:
                    datosextension = persona.datos_extension()
                    examenfisico = persona.datos_examen_fisico()
                    form = DatosMedicosForm(initial={
                        'carnetiess' : datosextension.carnetiess,
                        'sangre' : persona.sangre,
                        'archivotiposangre': persona.archivo_tiposangre ,
                        'peso' : examenfisico.peso,
                        'talla' : examenfisico.talla
                    })
                    data['form'] = form
                    template = get_template('th_hojavida/informacionmedica/modal/formbasico.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'editcontactoemergencia':
                try:
                    datosextension = persona.datos_extension()
                    form = ContactoEmergenciaForm(initial={
                        'contactoemergencia' : datosextension.contactoemergencia,
                        'parentescoemergencia' : datosextension.parentescoemergencia,
                        'telefonoemergencia' : datosextension.telefonoemergencia,
                        'telefonoconvemergencia': datosextension.telefonoconvemergencia,
                        'correoemergencia': datosextension.correoemergencia
                    })
                    data['form'] = form
                    template = get_template('th_hojavida/informacionmedica/modal/formbasico.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'addvacunacvd19':
                try:
                    data['formmodal'] = VacunacionCovidForm()
                    template = get_template("th_hojavida/modal/formvacunacion.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'adddosisvacuna':
                try:
                    data['filtro'] = filtro = VacunaCovid.objects.get(pk=int(request.GET['id']))
                    data['formmodal'] = VacunacionCovidForm()
                    template = get_template("th_hojavida/modal/formdosis.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'addcertificadovacunacovid':
                try:
                    data['filtro'] = filtro = VacunaCovid.objects.get(pk=int(request.GET['id']))
                    data['id']=request.GET['id']
                    data['form'] = VacunacionCovidEvidenciaForm()
                    template = get_template("ajaxformmodal.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            # CAPACITACIONES

            elif action == 'capacitaciones':
                try:
                    data['title'] = u'Capacitaciones'
                    url_vars, tipo = f'&action={action}', request.GET.get('tipo','')
                    url_vars += f'&tipo={tipo}'
                    if tipo == 'eventos':
                        listado = persona.distributivopersona_set.filter(status=True, estadopuesto=PUESTO_ACTIVO_ID)
                    elif tipo == 'solicitudes':
                        listado = persona.lista_evento_realizado_persona().order_by('-fechasolicitud')
                    elif tipo == 'perfeccionamiento':
                        listado = CapCabeceraSolicitudDocente.objects.filter(participante=persona, notificado=True, rutapdf__isnull=False).order_by('-fechasolicitud')
                    else:
                        listado = persona.mis_capacitaciones().order_by('-fechafin')
                    paging = MiPaginador(listado, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['url_vars'] = url_vars
                    data['tipo']=tipo
                    data['solicitud_capacitacion'] = variable_valor('SOLICITUD_CAPACITACION')
                    data['pendiente_capacitacion'] = variable_valor('PENDIENTE_CAPACITACION')
                    data['aprobado_capacitacion'] = variable_valor('APROBADO_CAPACITACION')
                    data['reporte_capacitaciones_persona'] = obtener_reporte('capacitaciones_persona')
                    data['listado'] = page.object_list
                    request.session['viewactivoth'] = ['informacionacademica', action]
                    return render(request, "th_hojavida/informacionacademica/capacitaciones.html", data)
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'loadEncuesta':
                from inno.models import CapPreguntaEncuestaPeriodo
                try:
                    id = encrypt(request.GET.get('id', encrypt('0')))
                    try:
                        eCapCabecera = CapCabeceraSolicitudDocente.objects.get(id=id)
                    except ObjectDoesNotExist:
                        raise NameError(u"No se encontro solicitud del evento")
                    eCapEvento = eCapCabecera.capeventoperiodo
                    eCapEncuesta = eCapEvento.encuesta
                    eCapPreguntas = CapPreguntaEncuestaPeriodo.objects.filter(encuesta=eCapEncuesta, status=True, isActivo=True)
                    data['eCapPreguntas'] = eCapPreguntas
                    data['frmName'] = "frmEncuesta"
                    data['eCapCabecera'] = eCapCabecera
                    template = get_template("th_hojavida/informacionacademica/modal/formencuestasatisfaccion.html")
                    json_content = template.render(data, request=request)
                    return JsonResponse({"isSuccess": True, 'html': json_content})
                except Exception as ex:
                    return JsonResponse({"isSuccess": False, 'message': str(ex)})

            elif action == 'detallecapacitacion':
                try:
                    data['capacitacion'] = capacitacion = Capacitacion.objects.get(pk=int(request.GET['id']))
                    if capacitacion.usuario_creacion:
                        data['personacreacion'] = Persona.objects.get(
                            usuario=capacitacion.usuario_creacion) if capacitacion.usuario_creacion.id > 1 else ""
                    template = get_template("th_hojavida/detallecapacitacion.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, "mensaje": f"Error al obtener los datos.{ex}"})

            elif action == 'detalleeventoaprobacion_docente':
                try:
                    data['cabecerasolicitud'] = cabecera = CapCabeceraSolicitudDocente.objects.get(
                        pk=int(request.GET['id']))
                    data['detallesolicitud'] = cabecera.capdetallesolicituddocente_set.all()
                    data['solicitud_capacitacion'] = variable_valor('SOLICITUD_CAPACITACION')
                    data['pendiente_capacitacion'] = variable_valor('PENDIENTE_CAPACITACION')
                    data['aprobado_capacitacion'] = variable_valor('APROBADO_CAPACITACION')
                    template = get_template("th_hojavida/detalleeventoaprobacion.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": True, 'data': json_content})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})


            elif action == 'addcapacitacion':
                try:
                    data['title'] = u'Adicionar capacitación'
                    form = CapacitacionPersonaForm()
                    form.adicionar()
                    visible_fields = form.visible_fields()
                    total_fields = len(visible_fields)
                    lista = [(1, 'Datos del evento', visible_fields[:8]),
                             (2, 'Datos de la capacitación', visible_fields[8:21]),
                             (3, 'Datos de ubicación', visible_fields[21:total_fields])
                             ]
                    data['form'] = lista
                    template = get_template('th_hojavida/informacionacademica/modal/formcapacitaciones.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'editcapacitacion':
                try:
                    data['title'] = u'Editar capacitación'
                    id=encrypt_id(request.GET['id'])
                    data['capacitacion'] = capacitacion = Capacitacion.objects.get(pk=id)
                    data['modalidad1'] = capacitacion.modalidad
                    form = CapacitacionPersonaForm(initial={'institucion': capacitacion.institucion,
                                                            'nombre': capacitacion.nombre,
                                                            'descripcion': capacitacion.descripcion,
                                                            'tipo': capacitacion.tipo,
                                                            'tipocurso': capacitacion.tipocurso,
                                                            'tipocertificacion': capacitacion.tipocertificacion,
                                                            'tipocapacitacion': capacitacion.tipocapacitacion,
                                                            'tipoparticipacion': capacitacion.tipoparticipacion,
                                                            'auspiciante': capacitacion.auspiciante,
                                                            'expositor': capacitacion.expositor,
                                                            'anio': capacitacion.anio,
                                                            'contexto': capacitacion.contextocapacitacion,
                                                            'detallecontexto': capacitacion.detallecontextocapacitacion,
                                                            'areaconocimiento': capacitacion.areaconocimiento,
                                                            'subareaconocimiento': capacitacion.subareaconocimiento,
                                                            'subareaespecificaconocimiento': capacitacion.subareaespecificaconocimiento,
                                                            'pais': capacitacion.pais,
                                                            'provincia': capacitacion.provincia,
                                                            'canton': capacitacion.canton,
                                                            'parroquia': capacitacion.parroquia,
                                                            'fechainicio': capacitacion.fechainicio,
                                                            'fechafin': capacitacion.fechafin,
                                                            'horas': capacitacion.horas,
                                                            'modalidad': capacitacion.modalidad,
                                                            'archivo': capacitacion.archivo,
                                                            'otramodalidad': capacitacion.otramodalidad}, instancia=capacitacion)
                    form.editar(capacitacion)
                    visible_fields = form.visible_fields()
                    total_fields = len(visible_fields)
                    lista = [(1, 'Datos del evento', visible_fields[:8]),
                             (2, 'Datos de la capacitación', visible_fields[8:21]),
                             (3, 'Datos de ubicación', visible_fields[21:total_fields])]
                    data['form'] = lista
                    data['id']=id
                    template = get_template('th_hojavida/informacionacademica/modal/formcapacitaciones.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'detalleevento':
                try:
                    data['evento'] = evento = CapEventoPeriodo.objects.get(pk=int(request.GET['id']))
                    template = get_template("th_hojavida/detalleevento.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": True, 'data': json_content})
                except Exception as ex:
                    return JsonResponse({"result": False, "mensaje": u"Error al obtener los datos."})

            elif action == 'detalleeventoaprobacion':
                try:
                    data['cabecerasolicitud'] = cabecera = CapCabeceraSolicitud.objects.get(pk=int(request.GET['id']))
                    data['detallesolicitud'] = cabecera.capdetallesolicitud_set.all()
                    data['solicitud_capacitacion'] = variable_valor('SOLICITUD_CAPACITACION')
                    data['pendiente_capacitacion'] = variable_valor('PENDIENTE_CAPACITACION')
                    data['aprobado_capacitacion'] = variable_valor('APROBADO_CAPACITACION')
                    template = get_template("th_hojavida/detalleeventoaprobacion.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": True, 'data': json_content})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": False, "mensaje": u"Error al obtener los datos."})

            elif action == 'detalleevento_docente':
                try:
                    data['evento'] = evento = CapEventoPeriodoDocente.objects.get(pk=int(request.GET['id']))
                    template = get_template("th_hojavida/detalleevento.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": True, 'data': json_content})
                except Exception as ex:
                    return JsonResponse({"result": False, "mensaje": u"Error al obtener los datos."})

            # CERTIFICACIONES

            elif action == 'certificaciones':
                try:
                    data['title'] = u'Certificaciones internacionales'
                    url_vars, tipo = f'&action={action}', request.GET.get('tipo','certificacionidiomas')
                    if tipo == 'certificacioninternacionales':
                        certificados = CertificacionPersona.objects.filter(status=True, persona=persona).order_by('-id')
                        data['reporte_certificaciones_persona'] = obtener_reporte('certificaciones_persona')
                    elif tipo == 'certificacionidiomas':
                        certificados = CertificadoIdioma.objects.filter(status=True, persona=persona).order_by('-id')
                        data['reporte_certificaciones_internacional'] = obtener_reporte('certificaciones_idiomas')
                    url_vars+=f'&tipo={tipo}'
                    paging = MiPaginador(certificados, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['url_vars'] = url_vars
                    data['t_certipersona']=len(CertificacionPersona.objects.filter(status=True, persona=persona).values_list('id'))
                    data['t_certiidioma']=len(CertificadoIdioma.objects.filter(status=True, persona=persona).values_list('id'))
                    data['tipo'] = tipo
                    data['listado'] = page.object_list
                    request.session['viewactivoth'] = ['informacionacademica', action]
                    return render(request, "th_hojavida/informacionacademica/certificaciones.html", data)
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'addcertificadoidioma':
                try:
                    data['persona'] = persona
                    form = CertificadoIdiomaForm()
                    data['switchery']=True
                    data['form'] = form
                    template = get_template('th_hojavida/informacionacademica/modal/formcertificadoidioma.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'editcertificadoidioma':
                try:
                    certificado = CertificadoIdioma.objects.get(pk=encrypt_id(request.GET['id']))
                    data['form'] = CertificadoIdiomaForm(initial=model_to_dict(certificado))
                    data['switchery'] = True
                    data['id']=certificado.id
                    template = get_template('th_hojavida/informacionacademica/modal/formcertificadoidioma.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'detalleotroscertificacion':
                try:
                    data['certificacion'] = certificacion = CertificadoIdioma.objects.get(pk=int(request.GET['id']))
                    if certificacion.usuario_creacion:
                        data['personacreacion'] = Persona.objects.get(
                            usuario=certificacion.usuario_creacion) if certificacion.usuario_creacion.id > 1 else ""
                    template = get_template("th_hojavida/detalleotroscertificacion.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, "mensaje": f"Error: {ex}"})

            elif action == 'addcertificado':
                try:
                    data['title'] = u'Adicionar Certificación'
                    data['persona'] = persona
                    form = CertificadoPersonaForm()
                    data['form'] = form
                    data['switchery'] = True
                    template = get_template('th_hojavida/informacionacademica/modal/formcertificado.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'editcertificado':
                try:
                    data['title'] = u'Editar Certificación'
                    data['certificado'] = certificado = CertificacionPersona.objects.get(pk=encrypt_id(request.GET['id']))
                    initial = model_to_dict(certificado)
                    form = CertificadoPersonaForm(initial=initial)
                    form.editar(certificado)
                    data['form'] = form
                    data['switchery'] = True
                    data['id'] = certificado.id
                    template = get_template('th_hojavida/informacionacademica/modal/formcertificado.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'detallecertificacion':
                try:
                    data['certificacion'] = certificacion = CertificacionPersona.objects.get(pk=int(request.GET['id']))
                    if certificacion.usuario_creacion:
                        data['personacreacion'] = Persona.objects.get(
                            usuario=certificacion.usuario_creacion) if certificacion.usuario_creacion.id > 1 else ""
                    template = get_template("th_hojavida/detallecertificacion.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, "mensaje": f"Error: {ex}"})

            # DISCAPACIDAD

            elif action == 'discapacidad':
                try:
                    data['title'] = u'Discapacidad'
                    data['subtitle'] = u'Información cargada'
                    data['url_vars'] = f'&action={action}'
                    data['persona'] = persona = Persona.objects.get(id=persona.id)
                    data['perfil'] = persona.mi_perfil()
                    data['documentopersonal'] = persona.documentos_personales()
                    request.session['viewactivoth'] = ['informacionmedica',action]
                    return render(request, "th_hojavida/informacionmedica/discapacidad.html", data)
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'editdiscapacidad':
                try:
                    from sga.models import Discapacidad
                    data['title'] = u'Discapacidad'
                    perfil = persona.mi_perfil()
                    form = DiscapacidadForm(initial=model_to_dict(perfil))
                    tienearchivo = True if perfil.archivo else False
                    data['form'] = form
                    data['tienearchivo'] = tienearchivo
                    data['switchery'] = True
                    data['tipodis'] = Discapacidad.objects.filter(status=True)
                    template = get_template('th_hojavida/informacionmedica/modal/formdiscapacidad.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            # DATOS LABORALES

            elif action == 'datosinstitucion':
                try:
                    data['title'] = u'Datos institucionales'
                    data['subtitle'] = u'Información cargada'
                    data['url_vars'] = f'&action={action}'
                    data['persona'] = persona = Persona.objects.get(id=persona.id)
                    data['perfil'] = persona.mi_perfil()
                    data['documentopersonal'] = persona.documentos_personales()
                    request.session['viewactivoth'] = ['informacionlaboral',action]
                    return render(request, "th_hojavida/informacionlaboral/datosinstitucion.html", data)
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'contratos':
                try:
                    data['title'] = u'Contratos'
                    url_vars=f'&action={action}'
                    contratos = persona.mis_contratos()
                    paging = MiPaginador(contratos, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['url_vars'] = url_vars
                    data['listado'] = page.object_list
                    request.session['viewactivoth'] = ['informacionlaboral',action]
                    return render(request, "th_hojavida/informacionlaboral/contratos.html", data)
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'rolpago':
                try:
                    data['title'] = u'Roles de pago'
                    url_vars=f'&action={action}'
                    roles = persona.rolpago_set.filter(periodo__estado=5, periodo__status=True, status=True)
                    paging = MiPaginador(roles, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['url_vars'] = url_vars
                    data['listado'] = page.object_list
                    data['reporte_0'] = obtener_reporte('rol_pago')

                    request.session['viewactivoth'] = ['informacionlaboral',action]
                    return render(request, "th_hojavida/informacionlaboral/rolpago.html", data)
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'detallerol':
                    try:
                        data['detallerol'] = registro = RolPago.objects.get(pk=int(request.GET['id']), status=True)
                        data['detalleinformativo'] = registro.detallerolinformativo()
                        data['detalleingreso'] = registro.detallerolingreso()
                        data['detalleegreso'] = registro.detallerolegreso_consolidado()
                        template = get_template("th_nomina/detallerol.html")
                        json_content = template.render(data)
                        return JsonResponse({"result": True, 'data': template.render(data)})
                    except Exception as ex:
                        messages.error(request, f'{ex}')

            elif action == 'evaluacion':
                try:
                    data['title'] = u'Evaluaciones'
                    url_vars, tipo = f'&action={action}', request.GET.get('tipo', '')

                    if not tipo:
                        listado = persona.podevaluaciondet_set.filter(status=True, podperiodo__publicacion__lte=datetime.now().date()).order_by( "podperiodo", "departamento")
                    elif tipo == 'docentepregrado':
                        url_vars += f'&tipo={tipo}'
                        data['existeanterior'] = ResumenFinalProcesoEvaluacionIntegral.objects.filter(profesor=profesor).exists()
                        data['periodoslect'] = MigracionEvaluacionDocente.objects.values('idperiodo', 'descperiodo', 'tipoeval').filter(idprofesor=profesor.id).distinct().order_by('idperiodo')
                        listado = RespuestaEvaluacionAcreditacion.objects.exclude(coordinacion=7).values('profesor', 'profesor__pk', 'proceso', 'proceso__pk',
                                                                                'proceso__mostrarresultados', 'proceso__periodo',
                                                                                'proceso__periodo__id',
                                                                                'proceso__periodo__nombre').filter(profesor=profesor.id).distinct().order_by('proceso')
                    elif tipo == 'docenteposgrado':
                        url_vars += f'&tipo={tipo}'
                        listado = RespuestaEvaluacionAcreditacion.objects.filter(proceso__mostrarresultados=False,coordinacion=7).values('profesor', 'proceso', 'proceso__pk',
                                                                     'proceso__mostrarresultados', 'proceso__periodo',
                                                                     'proceso__periodo__nombre').filter(profesor=profesor.id).distinct().order_by('proceso')
                    paging = MiPaginador(listado, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['tipo'] = tipo
                    data['url_vars'] = url_vars
                    data['listado'] = page.object_list
                    request.session['viewactivoth'] = ['informacionlaboral',action]
                    return render(request, "th_hojavida/informacionlaboral/evaluacion.html", data)
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'marcadas':
                try:
                    data['title'] = u'LOG de Marcadas'
                    # data['distributivo'] = distributivo = DistributivoPersona.objects.get(pk=int(encrypt(request.GET['id'])))
                    data['personaadminis'] = persona
                    data['anios'] = persona.lista_anios_trabajados_log()
                    data['jornadas'] = persona.historialjornadatrabajador_set.all()
                    puede_crear_marcada = False
                    try:
                        puede_crear_marcada = puede_realizar_accion_is_superuser(request, 'sagest.puede_crear_marcada')
                    except:
                        puede_crear_marcada = False
                    data['puede_crear_marcada'] = puede_crear_marcada
                    data['destino'] = 'th_hojavida'
                    data['pued_modificar'] = 1
                    # if not persona.id == persona.id:
                    #     raise NameError('Error')
                    data['hora'] = str(datetime.now().time())[0:5]
                    request.session['viewactivoth'] = ['informacionlaboral', action]
                    return render(request, "th_hojavida/informacionlaboral/marcadas.html", data)
                except Exception as ex:
                    pass

            #DECIMO
            elif action == 'decimo':
                try:
                    data['title'] = u'Décimo'
                    url_vars = f'&action={action}'
                    decimos = RegistroDecimo.objects.filter(status=True, persona=persona).order_by('-pk')

                    data['ultimo'] = ultimo = decimos.first()
                    paging = MiPaginador(decimos, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['configuracion'] = configuracion = ConfiguraDecimo.objects.filter(status=True).order_by('-pk').first()
                    data['url_vars'] = url_vars
                    data['listado'] = page.object_list
                    request.session['viewactivoth'] = ['informacionlaboral', action]
                    return render(request, "th_hojavida/informacionlaboral/decimo.html", data)
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'firmarcartadecimo':
                try:
                    data['id'] = encrypt_id(request.GET['id'])
                    # data['info_mensaje'] = f'Nota: Una vez seleccionado en firmar se firmara el certificado en la parte inferior del documneto.'
                    template = get_template("formfirmaelectronicamasiva.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')
                    return HttpResponseRedirect(request.path)

            elif action == 'descargarformato':
                try:
                    decimo = RegistroDecimo.objects.get(id=encrypt_id(request.GET['id']))
                    pdf, url=generar_carta_decimos(request, decimo)
                    return HttpResponseRedirect(url)
                except Exception as ex:
                    messages.error(request, f'{ex}')
                    return HttpResponseRedirect(request.path)

            # BITACORA

            elif action == 'mibitacora':
                try:
                    data['title'] = u'Bitácora '
                    url_vars, filtro, search, anio, mes = f'&action={action}', \
                                                          Q(status=True, persona=persona), \
                                                          request.GET.get('s', '').strip(), \
                                                          request.GET.get('anio', '').strip(), \
                                                          request.GET.get('mes', '').strip()
                    if anio:
                        filtro = filtro & Q(fecha__year=anio)
                        data['anio'] = int(anio)
                        url_vars += f'&anio={anio}'
                    if mes:
                        filtro = filtro & Q(fecha__month=mes)
                        data['mes'] = int(mes)
                        url_vars += f'&mes={mes}'
                    if search:
                        filtro = filtro & Q(Q(titulo__icontains=search) |
                                            Q(departamento__nombre__icontains=search) |
                                            Q(descripcion__icontains=search) |
                                            Q(link__icontains=search))
                        data['s'] = search
                        url_vars += f'&s={search}'
                    bitacoras = BitacoraActividadDiaria.objects.filter(filtro).order_by('-fecha')
                    paging = MiPaginador(bitacoras, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['bitacora_anios'] = BitacoraActividadDiaria.objects.filter(status=True, persona=persona).dates('fecha', 'year').distinct()
                    data['url_vars'] = url_vars
                    data['listado'] = page.object_list
                    data['reporte_bitacora'] = obtener_reporte('rpt_bitacora')
                    if persona.contratodip_set.filter(status=True).exists():
                        data['contratodip'] = persona.contratodip_set.filter(status=True)
                    puede_importar_actividades = False
                    if puede_realizar_accion_afirmativo(request, 'sga.puede_descargar_db_backup'):
                        puede_importar_actividades = True
                    data['puede_importar_actividades'] = puede_importar_actividades
                    request.session['viewactivoth'] = ['informacionlaboral', action]
                    return render(request, "th_hojavida/informacionlaboral/bitacora.html", data)
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'meses_anio_bitacora':
                try:
                    anio = request.GET['anio']
                    actividades=BitacoraActividadDiaria.objects.filter(persona=persona, fecha__year=anio,status=True).order_by('fecha').distinct()
                    lista = []
                    for e in actividades:
                        mes=MESES_CHOICES[e.fecha.month - 1][1]
                        if [e.fecha.month, mes.capitalize()] not in lista:
                            lista.append([e.fecha.month, mes.capitalize()])
                    return JsonResponse({"result": "ok", "data": lista})
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

            # EXPERIENCIA

            elif action == 'experiencia':
                try:
                    data['title'] = u'Experiencian laboral'
                    url_vars, filtro, search, = f'&action={action}', \
                                                          Q(status=True, persona=persona), \
                                                          request.GET.get('s', '').strip()
                    if search:
                        filtro = filtro & Q(Q(titulo__icontains=search))
                        data['s'] = search
                        url_vars += f'&s={search}'
                    experiencias = ExperienciaLaboral.objects.filter(filtro)
                    paging = MiPaginador(experiencias, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['url_vars'] = url_vars
                    data['listado'] = page.object_list
                    data['reporte_experiencias_persona'] = obtener_reporte('experiencias_persona')
                    request.session['viewactivoth'] = ['informacionlaboral', action]
                    return render(request, "th_hojavida/informacionlaboral/experiencialaboral.html", data)
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'addexperiencia':
                try:
                    data['title'] = u'Adicionar experiencia'
                    form = ExperienciaLaboralForm()
                    form.adicionar()
                    visible_fields = form.visible_fields()
                    total_fields = len(visible_fields)
                    lista = [(1, 'Información básica', visible_fields[:9]),
                             (2, 'Ubicación empresarial', visible_fields[9:13]),
                             (3, 'Datos laborales', visible_fields[13:total_fields])
                             ]
                    data['form'] = lista
                    data['switchery'] = True
                    template = get_template('th_hojavida/informacionlaboral/modal/formexperiencia.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'editexperiencia':
                try:
                    data['experiencia'] = experiencia = ExperienciaLaboral.objects.get(pk=encrypt_id(request.GET['id']))
                    data['id']=experiencia.id
                    vigente = True
                    fechafin = ""
                    if experiencia.motivosalida:
                        vigente = False
                        fechafin = experiencia.fechafin
                    form = ExperienciaLaboralForm(initial={'tipoinstitucion': experiencia.tipoinstitucion,
                                                                   'institucion': experiencia.institucion,
                                                                   'cargo': experiencia.cargo,
                                                                   'departamento': experiencia.departamento,
                                                                   'pais': experiencia.pais,
                                                                   'provincia': experiencia.provincia,
                                                                   'canton': experiencia.canton,
                                                                   'parroquia': experiencia.parroquia,
                                                                   'fechainicio': experiencia.fechainicio,
                                                                   'fechafin': fechafin,
                                                                   'motivosalida': experiencia.motivosalida,
                                                                   'vigente': vigente,
                                                                   'regimenlaboral': experiencia.regimenlaboral,
                                                                   'horassemanales': experiencia.horassemanales,
                                                                   'dedicacionlaboral': experiencia.dedicacionlaboral,
                                                                   'actividadlaboral': experiencia.actividadlaboral,
                                                                   'observaciones': experiencia.observaciones})
                    visible_fields = form.visible_fields()
                    total_fields = len(visible_fields)
                    lista = [(1, 'Información básica', visible_fields[:9]),
                             (2, 'Ubicación empresarial', visible_fields[9:13]),
                             (3, 'Datos laborales', visible_fields[13:total_fields])
                             ]
                    data['form'] = lista
                    data['switchery'] = True
                    template = get_template('th_hojavida/informacionlaboral/modal/formexperiencia.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            # MERITOS

            elif action == 'meritos':
                try:
                    data['title'] = u'Méritos'
                    url_vars, filtro, search, = f'&action={action}', \
                                                Q(status=True, persona=persona), \
                                                request.GET.get('s', '').strip()
                    if search:
                        filtro = filtro & Q(Q(titulo__icontains=search))
                        data['s'] = search
                        url_vars += f'&s={search}'
                    meritos = OtroMerito.objects.filter(filtro)
                    paging = MiPaginador(meritos, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['url_vars'] = url_vars
                    data['listado'] = page.object_list
                    request.session['viewactivoth'] = ['informacionacademica', action]
                    return render(request, "th_hojavida/informacionacademica/meritos.html", data)
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'addmerito':
                try:
                    # if perfilprincipal.es_estudiante:
                    data['persona'] = persona
                    form = OtroMeritoForm()
                    data['form'] = form
                    template = get_template('ajaxformmodal.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'editmerito':
                try:
                    # if perfilprincipal.es_estudiante:
                    data['title'] = u'Editar otros méritos'
                    data['merito'] = merito = OtroMerito.objects.get(pk=encrypt_id(request.GET['id']))
                    form = OtroMeritoForm( initial=model_to_dict(merito))
                    data['id']=merito.id
                    data['form'] = form
                    template = get_template('ajaxformmodal.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            # BECAS

            elif action == 'becas':
                try:
                    data['title'] = u'Becas'
                    url_vars, tipo = f'&action={action}', request.GET.get('tipo','')
                    if tipo == 'becasexternas':
                        filtro = Q(status=True, persona=persona)
                        becas = BecaPersona.objects.filter(filtro)
                    else:
                        filtro=Q(solicitud__inscripcion__persona=persona, status=True)
                        becas = BecaAsignacion.objects.filter(filtro)
                    paging = MiPaginador(becas, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['url_vars'] = url_vars
                    data['tipo'] = tipo
                    data['listado'] = page.object_list
                    request.session['viewactivoth'] = ['informacionacademica', action]
                    return render(request, "th_hojavida/informacionacademica/becas.html", data)
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'addbecaexterna':
                try:
                    data['title'] = u'Agregar Beca Externa'
                    form = DatosBecaForm()
                    # form.borrar_fecha_fin()
                    data['form'] = form
                    template = get_template('ajaxformmodal.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'editbecaexterna':
                try:
                    data['title'] = u'Editar Datos Beca Externa'
                    beca = BecaPersona.objects.get(pk=int(encrypt(request.GET['id'])))
                    form = DatosBecaForm(initial={'tipoinstitucion': beca.tipoinstitucion,
                                                  'institucion': beca.institucion,
                                                  'fechainicio': beca.fechainicio,
                                                  'fechafin': beca.fechafin}
                                         )

                    tienearchivo = True if beca.archivo else False

                    # if not beca.verificado:
                    #     form.borrar_fecha_fin()
                    # else:
                    #     form.bloquear_campos()

                    data['form'] = form
                    data['tienearchivo'] = tienearchivo
                    data['id'] = beca.id
                    template = get_template('ajaxformmodal.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            # INFORME DE ACTIVIDADES

            elif action == 'informesmensuales':
                try:
                    data['title'] = u'Informes mensuales'
                    url_vars = f'&action={action}'
                    informes=persona.informemensual_set.filter(status=True).order_by('-fechainicio')
                    paging = MiPaginador(informes, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['url_vars'] = url_vars
                    data['listado'] = page.object_list
                    request.session['viewactivoth'] = ['informacionlaboral', action]
                    return render(request, "th_hojavida/informacionlaboral/informes.html", data)
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'addinformemensual':
                try:
                    data['form'] = InformeMensualForm()
                    template = get_template('ajaxformmodal.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'editinformemensual':
                try:
                    data['informe'] = informe = persona.informemensual_set.get(pk=encrypt_id(request.GET['id']))
                    data['id']=informe.id
                    initial = model_to_dict(informe)
                    data['form'] = InformeMensualForm(initial=initial)
                    template = get_template('ajaxformmodal.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            # DEPORTISTA

            elif action == 'eventosdeportista':
                try:
                    data['title'] = u'Eventos de deporte'
                    url_vars = f'&action={action}'
                    eventos = DeportistaPersona.objects.filter(persona=persona, status=True).order_by('-id')
                    paging = MiPaginador(eventos, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['url_vars'] = url_vars
                    data['listado'] = page.object_list
                    request.session['viewactivoth'] = ['informacionpersonal', action]
                    return render(request, "th_hojavida/informacionacademica/eventosdeportista.html", data)
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'editdatosdeportista':
                try:
                    data['title'] = u'Editar Datos Deportista Alto Rendimiento'
                    deportista = DeportistaPersona.objects.get(pk=int(encrypt(request.GET['id'])))
                    form = DeportistaForm(initial={'disciplina': deportista.disciplina.all(),
                                                   'representapais': deportista.representapais,
                                                   'evento': deportista.evento,
                                                   'paisevento': deportista.paisevento,
                                                   'equiporepresenta': deportista.equiporepresenta,
                                                   'fechainicioevento': deportista.fechainicioevento,
                                                   'fechafinevento': deportista.fechafinevento,
                                                   'fechainicioentrena': deportista.fechainicioentrena,
                                                   'fechafinentrena': deportista.fechafinentrena,
                                                   'archivoentrena': deportista.archivoentrena,
                                                   'archivoevento': deportista.archivoevento,
                                                   'vigente': deportista.vigente
                                                   }
                                          )

                    tienearchivo = True if deportista.archivoevento else False

                    if deportista.verificado:
                        form.bloquear_campos()

                    data['form'] = form
                    data['id'] = deportista.id
                    template=get_template('th_hojavida/informacionacademica/modal/formdeportista.html')
                    return JsonResponse({'result':True, 'data':template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'adddatosdeportista':
                try:
                    data['title'] = u'Deportista Alto Rendimiento'
                    form = DeportistaForm()
                    form.borrar_vigente()
                    data['form'] = form
                    template = get_template('th_hojavida/informacionacademica/modal/formdeportista.html')
                    return JsonResponse({'result':True,'data':template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            # ARTISTAS

            elif action == 'artista':
                try:
                    data['title'] = u'Artista'
                    url_vars = f'&action={action}'
                    artistas = ArtistaPersona.objects.filter(persona=persona, status=True).order_by('-id')
                    paging = MiPaginador(artistas, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['url_vars'] = url_vars
                    data['listado'] = page.object_list
                    request.session['viewactivoth'] = ['informacionacademica', action]
                    return render(request, "th_hojavida/informacionacademica/artista.html", data)
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'adddatosartista':
                try:
                    data['title'] = u'Agregar Datos de Artista'

                    form = ArtistaForm()
                    form.borrar_vigente()

                    data['form'] = form
                    template=get_template("ajaxformmodal.html")
                    return JsonResponse({'result': True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'editdatosartista':
                try:
                    data['title'] = u'Editar Datos de Artista'

                    artista = ArtistaPersona.objects.get(pk=int(encrypt(request.GET['id'])))
                    form = ArtistaForm(initial={'campoartistico': artista.campoartistico.all(),
                                                'grupopertenece': artista.grupopertenece,
                                                'fechainicioensayo': artista.fechainicioensayo,
                                                'fechafinensayo': artista.fechafinensayo,
                                                'archivo': artista.archivo,
                                                'vigente': artista.vigente
                                                }
                                       )

                    tienearchivo = True if artista.archivo else False

                    if artista.verificado:
                        form.bloquear_campos()

                    data['form'] = form
                    data['id'] = artista.id
                    template = get_template("ajaxformmodal.html")
                    return JsonResponse({'result': True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            # ENFERMEDADES

            elif action == 'enfermedades':
                try:
                    data['title'] = u'Enfermedades'
                    url_vars = f'&action={action}'
                    artistas = persona.mis_enfermedades().order_by('-id')
                    paging = MiPaginador(artistas, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['url_vars'] = url_vars
                    data['listado'] = page.object_list
                    request.session['viewactivoth'] = ['informacionmedica', action]
                    return render(request, "th_hojavida/informacionmedica/enfermedades.html", data)
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'addenfermedad':
                try:
                    form = PersonaEnfermedadForm()
                    data['form'] = form
                    template = get_template('ajaxformmodal.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'editenfermedad':
                try:
                    data['enfermedad'] = enfermedad = PersonaEnfermedad.objects.get(
                        id=int(encrypt(request.GET['id'])))
                    data['id']=enfermedad.id
                    form = PersonaEnfermedadForm(initial=model_to_dict(enfermedad))
                    data['form'] = form
                    template = get_template('ajaxformmodal.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            # GASTOS PERSONALES

            elif action == 'gastospersonales':
                try:
                    data['title'] = u'Gastos personales'
                    url_vars = f'&action={action}'
                    gastospersonales = persona.gastospersonales_set.filter(status=True).order_by('-periodogastospersonales__anio', '-mes')
                    paging = MiPaginador(gastospersonales, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    formato = PeriodoGastosPersonales.objects.filter(status=True,mostrar=True).first()
                    if formato:
                        data['urlformato'] = formato.formato.url if formato.formato else ''
                    data['url_vars'] = url_vars
                    data['listado'] = page.object_list
                    request.session['viewactivoth'] = ['informacionpersonal', action]
                    return render(request, "th_hojavida/informacionpersonal/gastospersonales.html", data)
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'subirproyecciongastos':
                try:
                    data['filtro'] = GastosPersonales.objects.get(id=request.GET['id'])
                    template = get_template("th_hojavida/modal/formsubirdoc.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

            elif action == 'detallegasto':
                try:
                    data['gasto'] = gasto = GastosPersonales.objects.get(pk=encrypt_id(request.GET['id']), status=True)
                    data['empleado'] = persona = Persona.objects.get(pk=gasto.persona.id, status=True)
                    data['periodo'] = periodo = PeriodoGastosPersonales.objects.get(pk=gasto.periodogastospersonales.id)
                    data['detalles'] = ResumenMesGastosPersonales.objects.filter(gastospersonales__persona=persona, gastospersonales__periodogastospersonales__anio=periodo.anio).order_by('mes')
                    template = get_template("fin_gastospersonales/detallegasto.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": True, 'data': json_content})
                except Exception as ex:
                    return JsonResponse({"result": False, "mensaje": u"Error al obtener los datos."})

            # FORMULARIO 107

            elif action == 'formulario107':
                try:
                    data['title'] = u'Formulario 107'
                    url_vars = f'&action={action}'
                    formulario107 = persona.formulario107_set.filter(status=True).order_by('-anio')
                    paging = MiPaginador(formulario107, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['url_vars'] = url_vars
                    data['listado'] = page.object_list
                    request.session['viewactivoth'] = ['informacionpersonal', action]
                    return render(request, "th_hojavida/informacionpersonal/formulario107.html", data)
                except Exception as ex:
                    messages.error(request, f'{ex}')


            # SITUACIÓN LABORAL

            elif action == 'situacionlaboral':
                try:
                    data['title'] = u'Situación laboral'
                    url_vars = f'&action={action}'
                    historiallaboral = persona.personaaportacionhistoriallaboral_set.filter(status=True)
                    paging = MiPaginador(historiallaboral, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['url_vars'] = url_vars
                    data['listado'] = page.object_list
                    request.session['viewactivoth'] = ['informacionlaboral', action]
                    return render(request, "th_hojavida/informacionlaboral/situacionlaboral.html", data)
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'datossituacionlaboral':
                try:
                    data['title'] = u'Datos de Situacion Laboral'
                    data['action'] = action
                    form = DatosSituacionLaboralForm()
                    if persona.personasituacionlaboral_set.filter(status=True).exists():
                        data['situacion'] = situacion = PersonaSituacionLaboral.objects.get(persona=persona)
                        form = DatosSituacionLaboralForm(initial=model_to_dict(situacion))
                    data['form'] = form
                    data['switchery']=True
                    template=get_template("th_hojavida/datossituacionlaboral.html")
                    return JsonResponse({'result':True, 'data':template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'loadFormAportacionHistorialLaboral':
                try:
                    data['form'] = PersonaAportacionHistorialLaboralForm()
                    id = int(encrypt(request.GET['id'])) if request.GET['id'] else ''
                    data['frmName'] = 'frmAportacionHistorialLaboral'
                    if id:
                        data['registro'] = PersonaAportacionHistorialLaboral.objects.get(pk=id)
                        data['id'] = id
                    data['action']='saveAportacionHistorialLaboral'
                    template = get_template('ajaxformmodal.html')
                    json_content = template.render(data, request=request)
                    return JsonResponse({"result": True, 'data': json_content})
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al cargar los datos."})

            elif action == 'requisitosingreso':
                try:
                    data['title'] = u'Requisito periodo'
                    url_vars, filtro = f'&action={action}', Q(status=True)
                    listarequisitos = PersonaPeriodotthh.objects.filter(persona=persona,periodotthh__status=True, status=True)
                    paging = MiPaginador(listarequisitos, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['url_vars'] = url_vars
                    data['listado'] = page.object_list
                    request.session['viewactivoth'] = ['informacionlaboral', action]
                    return render(request, "th_hojavida/informacionlaboral/requisitoingreso.html", data)
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'subirrequisitosperiodotthh':
                try:
                    data['personaperiodotthh'] = personaperiodotthh = PersonaPeriodotthh.objects.filter(pk=encrypt_id(request.GET['id']), status=True).first()
                    data['id'] = personaperiodotthh.id
                    requisitos = RequisitoPeriodotthh.objects.filter(periodotthh_id=personaperiodotthh.periodotthh.id, mostrar=True, status=True)
                    if not personaperiodotthh.estado_requisito == 1:
                        for r in requisitos:
                            doc = r.documentopersonaperiodotthh_set.filter(personaperiodotthh=personaperiodotthh, status=True).first()
                            if not doc:
                                doc = DocumentoPersonaPeriodotthh(personaperiodotthh=personaperiodotthh, requisito=r, obligatorio = not r.opcional)
                                doc.save(request)
                                log(u'Adiciono documento requisito periodo: %s' % persona, request, "add")
                            else:
                                doc.opcional = not r.opcional
                                doc.save(request)
                    template = get_template("th_hojavida/informacionlaboral/modal/formrequisitosperiodo.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            # HOJA DE VIDA ANTERIOR
            elif action == 'hojavidaold':
                data['title'] = u'Hoja de vida'
                search = None
                data['datosextension'] = persona.datos_extension()
                data['documentopersonal'] = persona.documentos_personales()
                data['perfil'] = persona.mi_perfil()
                data['examenfisico'] = persona.datos_examen_fisico()
                data['reporte_0'] = obtener_reporte('rol_pago')
                data['reporte_1'] = obtener_reporte('hoja_vida_sagest')
                data['reporte_activos_persona'] = obtener_reporte('activos_persona')
                data['reporte_capacitaciones_persona'] = obtener_reporte('capacitaciones_persona')
                data['reporte_certificaciones_persona'] = obtener_reporte('certificaciones_persona')
                data['reporte_certificaciones_internacional'] = obtener_reporte('certificaciones_idiomas')
                data['reporte_experiencias_persona'] = obtener_reporte('experiencias_persona')
                data['reporte_publicaciones_profesor'] = obtener_reporte('publicaciones_profesor')
                data['reporte_bitacora'] = obtener_reporte('rpt_bitacora')
                data['perfilprincipal'] = perfilprincipal = request.session['perfilprincipal']
                # if Graduado.objects.filter(status=True,inscripcion__persona__id=persona.id).exists():
                #     f = Persona.objects.get(status=True, id=persona.id)
                #     data['datosactualizados'] = f.datosactualizados
                roles = persona.rolpago_set.filter(periodo__estado=5, periodo__status=True, status=True)
                paging = MiPaginador(roles, 25)
                p = 1
                try:
                    paginasesion = 1
                    if 'paginador' in request.session:
                        paginasesion = int(request.session['paginador'])
                    if 'page_rol' in request.GET:
                        p = int(request.GET['page_rol'])
                    else:
                        p = paginasesion
                    try:
                        page = paging.page(p)
                    except:
                        p = 1
                    page = paging.page(p)
                except:
                    page = paging.page(p)
                request.session['paginador'] = p
                data['paging_rol'] = paging
                data['rangospaging_rol'] = paging.rangos_paginado(p)
                data['page_rol'] = page
                if persona.tipocelular == 0:
                    data['tipocelular'] = '-'
                else:
                    data['tipocelular'] = TIPO_CELULAR[int(persona.tipocelular) - 1][1]
                data['roles'] = page.object_list
                # if 's' in request.GET:
                #     search = request.GET['s'].strip()
                #     ss = search.split(' ')
                #     if len(ss) == 1:
                #         activos = ActivoFijo.objects.filter(codigointerno=search, responsable=persona,
                #                                             statusactivo=1).order_by('ubicacion', 'descripcion')
                #     else:
                #         activos = ActivoFijo.objects.filter(responsable=persona, statusactivo=1).order_by('ubicacion',
                #                                                                                           'descripcion')
                # else:
                #     activos = ActivoFijo.objects.filter(responsable=persona, statusactivo=1).order_by('ubicacion',
                #                                                                                       'descripcion')
                # pagingactivo = MiPaginador(activos, 10)
                # p = 1
                # try:
                #     paginasesion = 1
                #     if 'paginador' in request.session:
                #         paginasesion = int(request.session['paginador'])
                #     else:
                #         p = paginasesion
                #     if 'page_activo' in request.GET:
                #         p = int(request.GET['page_activo'])
                #     else:
                #         p = paginasesion
                #     try:
                #         page = pagingactivo.page(p)
                #     except:
                #         p = 1
                #     page = pagingactivo.page(p)
                # except:
                #     page = paging.page(p)
                # request.session['paginador'] = p
                # data['paging_activo'] = pagingactivo
                data['administrativo'] = persona
                id = int(persona.id)
                puede_importar_actividades = False
                # if str(id) in variable_valor('PUEDE_IMPORTAR_ACTIVIDADES_BITACORA'):
                if puede_realizar_accion_afirmativo(request, 'sga.puede_descargar_db_backup'):
                    puede_importar_actividades = True
                data['puede_importar_actividades'] = puede_importar_actividades
                persona2 = Persona.objects.get(pk=id)
                correo = persona2.email
                data['correo'] = correo
                data['fechanacimiento'] = persona2.nacimiento
                data['persona2'] = persona2
                # data['search'] = search if search else ""
                # data['rangospaging_activo'] = pagingactivo.rangos_paginado(p)
                # data['page_activo'] = page
                # data['activos'] = page.object_list
                data['gastospersonales'] = persona.gastospersonales_set.all().order_by('-periodogastospersonales__anio',
                                                                                       '-mes')
                data['pod'] = persona.podevaluaciondet_set.filter(status=True,
                                                                  podperiodo__publicacion__lte=datetime.now().date()).order_by(
                    "podperiodo", "departamento")
                data['puede_ingresar_gastos'] = PeriodoGastosPersonales.objects.filter(
                    fechadesde__lte=datetime.now().date(), fechahasta__gte=datetime.now().date()).exists()
                if persona.es_administrativo() or persona.es_profesor():
                    data['articulos'] = ArticuloInvestigacion.objects.select_related().filter((Q(
                        participantesarticulos__profesor__persona=persona) | Q(
                        participantesarticulos__administrativo__persona=persona) | Q(
                        participantesarticulos__inscripcion__persona=persona)), status=True, aprobado=True,
                                                                                              participantesarticulos__status=True).order_by(
                        '-fechapublicacion')
                    data['ponencias'] = PonenciasInvestigacion.objects.select_related().filter((Q(
                        participanteponencias__profesor__persona=persona) | Q(
                        participanteponencias__administrativo__persona=persona) | Q(
                        participanteponencias__inscripcion__persona=persona)), status=True,
                                                                                               participanteponencias__status=True)
                    data['capitulolibro'] = CapituloLibroInvestigacion.objects.select_related().filter((Q(
                        participantecapitulolibros__profesor__persona=persona) | Q(
                        participantecapitulolibros__profesor__persona=persona)), status=True,
                                                                                                       participantecapitulolibros__status=True)
                    data['libros'] = LibroInvestigacion.objects.select_related().filter((Q(
                        participantelibros__profesor__persona=persona) | Q(participantelibros__profesor__persona=persona)),
                                                                                        status=True,
                                                                                        participantelibros__status=True)
                    data['solicitudes'] = SolicitudPublicacion.objects.filter(persona=persona, aprobado=False,
                                                                              status=True).order_by('-fecha_creacion')
                    data['revisiones'] = ParRevisorProduccionCientifica.objects.filter(persona=persona,
                                                                                       status=True).order_by(
                        '-fecharevision')

                    data['redes'] = RedPersona.objects.filter(persona=persona, status=True)
                    bitacoras = BitacoraActividadDiaria.objects.filter(persona=persona, status=True).order_by('-fecha')
                    bitacoraspos = BitacoraActividadDiaria.objects.filter(persona=persona, status=True).order_by('-fecha')
                    searchb = None
                    searchbp = None
                    if 'sb' in request.GET:
                        searchb = request.GET['sb'].strip()
                        bitacoras = bitacoras.filter(Q(titulo__icontains=searchb) |
                                                     Q(departamento__nombre__icontains=searchb) |
                                                     Q(descripcion__icontains=searchb) |
                                                     Q(link__icontains=searchb))
                    if 'sbp' in request.GET:
                        searchbp = request.GET['sbp'].strip()
                        bitacoraspos = bitacoraspos.filter(Q(titulo__icontains=searchbp) |
                                                           Q(departamento__nombre__icontains=searchbp) |
                                                           Q(descripcion__icontains=searchbp) |
                                                           Q(link__icontains=searchbp))
                    aniobit = None
                    aniobitpos = None
                    if 'aniobit' in request.GET:
                        if request.GET['aniobit']:
                            aniobit = request.GET['aniobit'].strip()
                            bitacoras = bitacoras.filter(fecha__year=aniobit)

                    if 'aniobitpos' in request.GET:
                        if request.GET['aniobitpos']:
                            aniobitpos = request.GET['aniobitpos'].strip()
                            bitacoraspos = bitacoraspos.filter(fecha__year=aniobitpos)

                    mesbit = None
                    mesbitpos = None
                    if 'mesbit' in request.GET:
                        if request.GET['mesbit']:
                            mesbit = request.GET['mesbit'].strip()
                            bitacoras = bitacoras.filter(fecha__month=mesbit)

                    pagingbitacora = MiPaginador(bitacoras, 10)
                    pagingbitacorapos = MiPaginador(bitacoraspos, 10)
                    p = 1
                    ppos = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        else:
                            p = paginasesion
                        if 'page_bitacora' in request.GET:
                            p = int(request.GET['page_bitacora'])
                        else:
                            p = paginasesion
                        try:
                            page = pagingbitacora.page(p)
                        except:
                            p = 1
                        page = pagingbitacora.page(p)
                    except:
                        page = paging.page(p)
                    # try:
                    #     paginasesionpos = 1
                    #     if 'paginadorpos' in request.session:
                    #         paginasesionpos = int(request.session['paginadorpos'])
                    #     else:
                    #         ppos = paginasesionpos
                    #     if 'page_bitacora_pos' in request.GET:
                    #         ppos = int(request.GET['page_bitacora_pos'])
                    #     else:
                    #         ppos = paginasesionpos
                    #     try:
                    #         pagepos = pagingbitacorapos.page(ppos)
                    #     except:
                    #         ppos = 1
                    #     pagepos = pagingbitacorapos.page(ppos)
                    # except:
                    #     pagepos = paging.page(p)
                    request.session['paginador'] = p
                    data['paging_bitacora'] = pagingbitacora
                    data['paging_bitacora_pos'] = pagingbitacorapos
                    data['searchb'] = searchb if searchb else ""
                    data['aniobit'] = int(aniobit) if aniobit else ""
                    data['mesbit'] = int(mesbit) if mesbit else ""
                    data['rangospaging_bitacora'] = pagingbitacora.rangos_paginado(p)
                    data['page_bitacora'] = page
                    data['bitacoras'] = page.object_list
                    data['bitacora_anios'] = bitacoras.dates('fecha', 'year').distinct()

                    # data['searchbp'] = searchbp if searchbp else ""
                    # data['aniobitpos'] = int(aniobitpos) if aniobitpos else ""
                    # data['mesbitpos'] = int(mesbitpos) if mesbitpos else ""
                    # data['rangospaging_bitacora_pos'] = pagingbitacorapos.rangos_paginado(ppos)
                    # data['page_bitacora_pos'] = pagepos
                    # data['bitacoras_pos'] = pagepos.object_list
                    # data['bitacora_anios_pos'] = bitacoraspos.dates('fecha', 'year').distinct()

                    distributivos = persona.distributivopersona_set.all()
                    data['anios'] = persona.lista_anios_trabajados_log()
                    data['jornadas'] = persona.historialjornadatrabajador_set.all()
                    if distributivos:
                        data['distributivo'] = distributivo = distributivos[0]
                        # data['anios'] = distributivo.persona.lista_anios_trabajados()
                        # data['jornadas'] = distributivo.persona.historialjornadatrabajador_set.all()
                    else:
                        data['distributivo'] = None
                        # data['anios'] = None
                        # data['jornadas'] = None
                    data['listacargo'] = persona.distributivopersona_set.filter(status=True, estadopuesto=1)
                    data['solicitud_capacitacion'] = variable_valor('SOLICITUD_CAPACITACION')
                    data['pendiente_capacitacion'] = variable_valor('PENDIENTE_CAPACITACION')
                    data['aprobado_capacitacion'] = variable_valor('APROBADO_CAPACITACION')
                    data['formularios107'] = formulario = Formulario107.objects.filter(status=True,
                                                                                       persona=persona).order_by('-anio')
                    if formulario:
                        data['anioselect'] = formulario[0].id
                    else:
                        data['anioselect'] = 0
                else:
                    data['distributivo'] = None
                    data['anios'] = None
                    data['jornadas'] = None
                # if persona.es_estudiante():
                #     data['inscripcion'] = perfilprincipal.inscripcion

                if perfilprincipal.es_estudiante():
                    data['archivos'] = perfilprincipal.inscripcion.archivo_set.filter(status=True)
                    data['inscripcion'] = perfilprincipal.inscripcion
                # listadocentes = ParticipantesMatrices.objects.values('proyecto__programa__nombre', 'proyecto__nombre', 'proyecto__tipo', 'horas', 'tipoparticipante__nombre').filter(matrizevidencia_id=2, status=True, proyecto__status=True, profesor__persona=persona)
                listadocentes = None
                if perfilprincipal.es_profesor():
                    if persona.profesor_set.filter(status=True).exists():
                        profesor = persona.profesor()
                        listadocentes = profesor.participantesmatrices_set.filter(status=True, proyecto__status=True)
                # listadocentes = ParticipantesMatrices.objects.filter(matrizevidencia_id=2, status=True, proyecto__status=True, profesor__persona=persona)
                # listaestudiante = ParticipantesMatrices.objects.values('proyecto__programa__nombre', 'proyecto__nombre', 'proyecto__tipo', 'horas', 'tipoparticipante__nombre', 'inscripcion__carrera__nombre', 'actividad__titulo').filter(matrizevidencia_id=2, status=True, proyecto__status=True, inscripcion__persona=persona)
                proyectosnoconvalidados = ParticipantesMatrices.objects.values('id', 'proyecto__programa__nombre',
                                                                               'proyecto__nombre', 'proyecto__tipo',
                                                                               'horas', 'tipoparticipante__nombre',
                                                                               'inscripcion__carrera__nombre',
                                                                               'actividad__titulo').filter(
                    matrizevidencia_id=2, status=True, proyecto__status=True, inscripcion__persona=persona,
                    actividad__isnull=True)
                proyectosconvalidados = ParticipantesMatrices.objects.values('id', 'proyecto__programa__nombre',
                                                                             'proyecto__nombre', 'proyecto__tipo', 'horas',
                                                                             'tipoparticipante__nombre',
                                                                             'inscripcion__carrera__nombre',
                                                                             'actividad__titulo').filter(
                    matrizevidencia_id=2, status=True, inscripcion__persona=persona, actividad__isnull=False,
                    proyecto__status=True)

                if proyectosconvalidados and proyectosnoconvalidados:
                    listaestudiante = proyectosnoconvalidados | proyectosconvalidados
                elif proyectosnoconvalidados:
                    listaestudiante = proyectosnoconvalidados
                else:
                    listaestudiante = proyectosconvalidados

                if listaestudiante:
                    listaestudiante = sorted(listaestudiante, key=lambda i: i['id'], reverse=True)

                # listaadministrativo = ParticipantesMatrices.objects.values('proyecto__programa__nombre', 'proyecto__nombre', 'proyecto__tipo', 'horas', 'tipoparticipante__nombre').filter(matrizevidencia_id=2, status=True, proyecto__status=True, administrativo__persona=persona)
                # listaadministrativo = ParticipantesMatrices.objects.filter(matrizevidencia_id=2, status=True, proyecto__status=True, administrativo__persona=persona)
                if persona.administrativo_set.filter(status=True):
                    listaadministrativo = persona.administrativo_set.filter(
                        status=True).first().participantesmatrices_set.filter(status=True, proyecto__status=True)
                else:
                    listaadministrativo = None
                # listaadministrativo = ParticipantesMatrices.objects.filter(matrizevidencia_id=2, status=True, proyecto__status=True, administrativo__persona=persona)
                if listadocentes and listaadministrativo:
                    data['proyectos'] = listadocentes | listaadministrativo
                elif listaadministrativo:
                    data['proyectos'] = listaadministrativo
                elif listadocentes:
                    data['proyectos'] = listadocentes
                data['proyectos_alu'] = listaestudiante
                data['proyectos_externo'] = ProyectoInvestigacionExterno.objects.filter(persona=persona, status=True)
                data['reporte_2'] = obtener_reporte('certificado_ficha_estudiantil')
                # data['reporte_3'] = obtener_reporte('certificado_laboral')
                # data['niveltitulo'] = NivelTitulacion.objects.filter(status=True).order_by('-rango')
                idtitulaciones = Titulacion.objects.filter(status=True, persona=persona).values_list('titulo__nivel__id',
                                                                                                     flat=True)
                data['niveltitulo'] = NivelTitulacion.objects.filter(status=True, id__in=idtitulaciones)
                data['listacargo'] = persona.distributivopersona_set.filter(status=True, estadopuesto=PUESTO_ACTIVO_ID)
                data['es_profesor'] = es_profesor = False
                if perfilprincipal.es_profesor():
                    if persona.profesor_set.filter(status=True).exists():
                        data['es_profesor'] = es_profesor = True
                        profesor = persona.profesor().id
                        data['profesor'] = profesor
                        data['titulacionpostgrado'] = titulacionpostgrado = TemaTitulacionPosgradoMatricula.objects.filter(
                            status=True, tutor=profesor).order_by('pk')
                        data['existeactual'] = evaluaciondocente = RespuestaEvaluacionAcreditacion.objects.filter(
                            proceso__mostrarresultados=True).values('profesor', 'profesor__pk', 'proceso', 'proceso__pk',
                                                                    'proceso__mostrarresultados', 'proceso__periodo',
                                                                    'proceso__periodo__id',
                                                                    'proceso__periodo__nombre').filter(
                            profesor=profesor).distinct().order_by('proceso')
                        data[
                            'existeactualposgrado'] = evaluaciondocenteposgrado = RespuestaEvaluacionAcreditacion.objects.filter(
                            proceso__mostrarresultados=False).values('profesor', 'proceso', 'proceso__pk',
                                                                     'proceso__mostrarresultados', 'proceso__periodo',
                                                                     'proceso__periodo__nombre').filter(
                            profesor=profesor).distinct().order_by('proceso')
                        data['existe'] = RespuestaEvaluacionAcreditacion.objects.filter(profesor=profesor,
                                                                                        tipoinstrumento=1,
                                                                                        proceso__mostrarresultados=True).exists()
                        data['existeposgrado'] = RespuestaEvaluacionAcreditacion.objects.filter(profesor=profesor,
                                                                                                tipoinstrumento=1,
                                                                                                proceso__mostrarresultados=False).exists()
                        data['existeanterior'] = ResumenFinalProcesoEvaluacionIntegral.objects.filter(
                            profesor=profesor).exists()
                        data['periodoslect'] = MigracionEvaluacionDocente.objects.values('idperiodo', 'descperiodo',
                                                                                         'tipoeval').filter(
                            idprofesor=profesor).distinct().order_by('idperiodo')
                        data['periodoslect'] = MigracionEvaluacionDocente.objects.values('idperiodo', 'descperiodo',
                                                                                         'tipoeval').filter(
                            idprofesor=profesor).distinct().order_by('idperiodo')

                data['otromeritos'] = OtroMerito.objects.filter(status=True, persona=persona)
                data['referencias'] = ReferenciaPersona.objects.filter(status=True, persona=persona)
                data['certificaciones'] = CertificacionPersona.objects.filter(status=True, persona=persona)

                data['otrascertificaciones'] = CertificadoIdioma.objects.filter(status=True, persona=persona).order_by('id')
                data['certificadostutorias'] = CertificadoTutoriaHV.objects.filter(status=True, persona=persona).order_by(
                    'id')

                data['informesmensuales'] = persona.informemensual_set.filter(status=True).order_by('-fechainicio')

                lista_becas = []
                becasinternas = BecaAsignacion.objects.filter(solicitud__inscripcion__persona=persona, status=True)
                for beca in becasinternas:
                    lista_becas.append(['INTERNA', 'PÚBLICA', 'UNIVERSIDAD ESTATAL DE MILAGRO', beca.solicitud.periodo,
                                        beca.solicitud.becatipo, None, beca.fecha,
                                        'NO' if beca.solicitud.periodo.finalizo() else 'SI', None, None, None, None])

                becasexternas = BecaPersona.objects.filter(persona=persona, status=True)
                for beca in becasexternas:
                    becaarchivo = None
                    becafechainicio = None
                    if beca.archivo:
                        becaarchivo = beca.archivo.url
                    if beca.fechainicio:
                        becafechainicio = beca.fechainicio
                    lista_becas.append(
                        ['EXTERNA', beca.get_tipoinstitucion_display(), beca.institucion.nombre, None, None, becaarchivo,
                         becafechainicio, 'NO' if beca.fechafin else 'SI', beca.estadoarchivo,
                         beca.get_estadoarchivo_display(), beca.observacion, encrypt(beca.id)])

                lista_becas.sort(key=lambda beca: beca[6], reverse=True)

                data['becas'] = lista_becas
                data['deportista'] = DeportistaPersona.objects.filter(persona=persona, status=True).order_by('-id')
                data['migrante'] = MigrantePersona.objects.filter(persona=persona)[0] if MigrantePersona.objects.filter(
                    persona=persona).exists() else None
                data['artistas'] = ArtistaPersona.objects.filter(persona=persona, status=True).order_by('-id')
                data['vacunascovid'] = VacunaCovid.objects.filter(persona=persona, status=True).order_by('-id')
                banderadeclaracion = 0
                mensaje = ''
                d = DeclaracionBienes.objects.filter(persona=persona, fechaproximoregistro__isnull=False,
                                                     status=True).order_by('-pk').first()
                tiene = DeclaracionBienes.objects.filter(persona=persona, status=True)
                if d:
                    fechaproximoregistro = d.fechaproximoregistro
                    if fechaproximoregistro:
                        if datetime.now().date() > fechaproximoregistro:
                            banderadeclaracion = 1
                            fechadeclaracion = d.fecha
                            if fechadeclaracion:
                                mensaje = u'Tiene pendiente cargar el documento de la declaración, ' \
                                          u'el último registro tiene fecha %s y su próxima declaración debe ' \
                                          u'ser en: %s' % (str(fechadeclaracion), str(fechaproximoregistro))
                            else:
                                mensaje = u'Tiene pendiente cargar el documento de la declaración'
                else:
                    if not tiene:
                        banderadeclaracion = 1
                        mensaje = u'Tiene pendiente cargar el documento de la declaración'
                    else:
                        banderadeclaracion = 0
                        mensaje = ''
                if esestudiante:
                    banderadeclaracion = 0
                    mensaje = ''

                data['banderadeclaracion'] = banderadeclaracion
                formato = PeriodoGastosPersonales.objects.filter(status=True, mostrar=True).first()
                if formato:
                    data['urlformato'] = formato.formato.url if formato.formato else ''
                data['aportacioneshistoriallaboral'] = persona.personaaportacionhistoriallaboral_set.filter(status=True)
                data['mensaje'] = mensaje
                data['puede_modificar_hv'] = variable_valor('PUEDE_MODIFICAR_HV')
                data['ficha'] = persona.mi_ficha()
                if 'tab' in request.GET:
                    data['tab'] = request.GET['tab']

                try:
                    if es_profesor:
                        # data['actualizarlenguamaterna'] = actualizarlenguamaterna = persona.idiomadomina_set.filter(status=True, lenguamaterna=True).exists()
                        data['existepromocion'] = existepromocion = PeriodoPromocionDocente.objects.filter(status=True,
                                                                                                           fechafin__gte=datetime.now().date()).exists()
                        if existepromocion:
                            data['periodopromocion'] = periodopromocion = PeriodoPromocionDocente.objects.filter(
                                status=True, fechafin__gte=datetime.now().date()).first()
                    if persona.contratodip_set.filter(status=True).exists():
                        data['contratodip'] = persona.contratodip_set.filter(status=True)
                    # data['solicitudestraspasosat'] = SolicitudActivos.objects.filter(Q(status=True) & ((Q(responsableasignacion=persona) & Q(estado=4)) | (Q(activo__responsable=persona) & Q(estado=1))))
                    # data['solicitudmantenimiento'] = SolicitudConfirmacionMantenimiento.objects.filter(status=True,mantenimiento__activotecno__activotecnologico__responsable=persona)
                    if perfilprincipal.es_inscripcionaspirante():
                        return render(request, "th_hojavida/viewadmisionposgrado.html", data)
                    else:
                        return render(request, "th_hojavida/view.html", data)
                except Exception as ex:
                    # print(ex)
                    return HttpResponseRedirect(f"/?info={ex.__str__()}")

            # INFORMES DE ACTIVIDADES POSGRADO
            elif action == 'informesposgrado':
                try:
                    data['title'] = u'Informes mensuales posgrado'
                    url_vars = f'&action={action}'
                    query = SolicitudPago.objects.filter(status=True,contrato__persona=persona)
                    paging = MiPaginador(query, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['url_vars'] = url_vars
                    data['listado'] = page.object_list
                    return render(request,'th_hojavida/informacionlaboral/informesposgrado.html',data)
                except Exception as ex:
                    pass

            elif action == 'viewrequisitosposgrado':
                try:
                    id = encrypt(request.GET['id'])
                    registro = RequisitoSolicitudPago.objects.filter(status=True,solicitud_id=int(id))
                    data['lista'] = registro
                    template = get_template(
                        'th_hojavida/informacionlaboral/modal/../templates/pro_solicitudpago/modal/viewrequisitosposgrado.html')
                    res_js={'result':True,'data':template.render(data)}
                except Exception as ex:
                    err_ = f'Ocurrio un error: {ex.__str__()}. En la linea {sys.exc_info()[-1].tb_lineno}'
                    res_js={'result':False,'mensaje':err_}
                return JsonResponse(res_js)

            elif action == 'viewhistorialrequisito':
                try:
                    id = encrypt(request.GET['id'])
                    reg = HistorialProcesoSolicitud.objects.filter(status=True,requisito_id=int(id))
                    data['lista'] = reg
                    template = get_template('th_hojavida/informacionlaboral/modal/viewhistoryrequiposgrado.html')
                    res_js = {'result':True,'data':template.render(data)}
                except Exception as ex:
                    err_ = f'Ocurrio un error: {ex.__str__()}. En la linea {sys.exc_info()[-1].tb_lineno}'
                    res_js = {'result':False,'mensaje':err_}
                return JsonResponse(res_js)

            # PAZ Y SALVO
            elif action == 'pazsalvo':
                try:
                    data['title'] = u'Paz y salvo'
                    url_vars, filtro = f'&action={action}', Q(status=True, persona=persona)
                    pazsalvos = PazSalvo.objects.filter(filtro).order_by('-id')
                    paging = MiPaginador(pazsalvos, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['url_vars'] = url_vars
                    data['existe_requisitos'] = RequisitoPazSalvo.objects.filter(status=True, mostrar=True)
                    data['listado'] = page.object_list
                    request.session['viewactivoth'] = ['informacionlaboral', action]
                    return render(request, "th_hojavida/informacionlaboral/pazsalvo.html", data)
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'adddecimo':
                try:
                    form = RegistraDecimoForm()
                    data['form'] = form
                    template = get_template("ajaxformmodal.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'addpazsalvo':
                try:
                    form = PazSalvoFormHV()
                    ids_d = DistributivoPersona.objects.filter(status=True, persona=persona).values_list('unidadorganica_id', flat=True).distinct()
                    ids_h = DistributivoPersonaHistorial.objects.filter(status=True, persona=persona).values_list('unidadorganica_id', flat=True).distinct()
                    ids_departamentos = list(ids_d) + list(ids_h)
                    form.fields['jefeinmediato'].queryset = Persona.objects.none()
                    form.fields['cargo'].queryset = DenominacionPuesto.objects.none()
                    form.fields['departamento'].queryset = Departamento.objects.filter(id__in=ids_departamentos)
                    data['form'] = form
                    template = get_template("th_hojavida/modal/formpazsalvo.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'editpazsalvo':
                try:
                    data['id'] = id = encrypt_id(request.GET['id'])
                    pazsalvo=PazSalvo.objects.get(id=id)
                    form = PazSalvoFormHV(model_to_dict(pazsalvo))
                    ids_d = DistributivoPersona.objects.filter(status=True,unidadorganica__status=True, persona=persona).values_list('unidadorganica_id', flat=True).distinct()
                    ids_h = DistributivoPersonaHistorial.objects.filter(status=True,unidadorganica__status=True, persona=persona).values_list('unidadorganica_id', flat=True).distinct()
                    ids_departamentos = list(ids_d) + list(ids_h)
                    form.fields['jefeinmediato'].queryset = Persona.objects.filter(id=pazsalvo.jefeinmediato.id)
                    form.fields['cargo'].queryset = DenominacionPuesto.objects.filter(id=pazsalvo.cargo.id)
                    form.fields['departamento'].queryset = Departamento.objects.filter(id__in=ids_departamentos)
                    data['form'] = form
                    template = get_template("th_hojavida/modal/formpazsalvo.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'cargarcargos':
                try:
                    id = request.GET.get('id', '')
                    ids_p = DistributivoPersona.objects.filter(status=True, unidadorganica_id=id, persona=persona).values_list('denominacionpuesto_id', flat=True).distinct()
                    ids_ph = DistributivoPersonaHistorial.objects.filter(status=True, persona=persona, unidadorganica_id=id).values_list('denominacionpuesto_id', flat=True).distinct()
                    ids_cargo = list(ids_p) + list(ids_ph)
                    cargos = DenominacionPuesto.objects.filter(id__in=ids_cargo)
                    resp = [{'value': c.pk, 'text': f"{c.descripcion}"} for c in cargos]
                    return JsonResponse({'result': True, 'data': resp})
                except Exception as ex:
                    pass

            elif action == 'consultarcargos':
                try:
                    id = request.GET.get('id', '')
                    tipodeclaracion = request.GET.get('args', '')
                    filtro = Q(status=True)
                    if id == '1':
                        ids_p = DistributivoPersona.objects.filter(status=True, persona=persona).values_list('denominacionpuesto_id', flat=True).distinct()
                        filtro = filtro & Q(id__in=ids_p)
                    elif id == '0':
                        if tipodeclaracion == '2' or tipodeclaracion == '3':
                            ids_p = DistributivoPersonaHistorial.objects.filter(status=True, persona=persona).values_list('denominacionpuesto_id', flat=True).distinct()
                            filtro = filtro & Q(id__in=ids_p)
                    cargos = DenominacionPuesto.objects.filter(filtro).order_by('descripcion')
                    resp = [{'value': c.pk, 'text': f"{c.descripcion}"} for c in cargos]
                    return JsonResponse({'result': True, 'data': resp})
                except Exception as ex:
                    pass

            elif action == 'cargarrmu':
                try:
                    idcargo = request.GET.get('value', '')
                    cargo = DenominacionPuesto.objects.get(id=idcargo)
                    distributivo = DistributivoPersona.objects.filter(status=True, denominacionpuesto=cargo, persona=persona).order_by('-fecha_creacion').first()
                    if not distributivo:
                        distributivo = DistributivoPersonaHistorial.objects.filter(status=True, denominacionpuesto=cargo, persona=persona).order_by('-fechahistorial').first()
                    rmu = distributivo.rmupuesto if distributivo else 0
                    return JsonResponse({'result': True, 'rmu': rmu})
                except Exception as ex:
                    pass

            elif action == 'firmarpazsalvo':
                try:
                    data['id'] = encrypt_id(request.GET['id'])
                    # data['info_mensaje'] = f'Nota: Una vez seleccionado en firmar se firmara el certificado en la parte inferior del documneto.'
                    template = get_template("formfirmaelectronicamasiva.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')
                    return HttpResponseRedirect(request.path)

            elif action == 'firmarpazsalvo_new':
                try:
                    data['id'] = id = encrypt_id(request.GET['id'])
                    qr = qrImgFirma(request, persona, "png", paraMostrar=True)
                    data["qrBase64"] = qr[0]
                    pazsalvo = PazSalvo.objects.get(id=id)
                    certificado_firma = pazsalvo.documento()
                    archivo_ = certificado_firma.ultimo_archivo_con_exclusion(persona).archivo
                    data['filtro'] = pazsalvo
                    data['archivo_url'] = archivo_.url
                    template = get_template("formfirmaelectronica_v2.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'subirrequisitos':
                try:
                    data['pazsalvo'] = pazsalvo = PazSalvo.objects.get(id=encrypt_id(request.GET['id']))
                    data['id'] = pazsalvo.id
                    requisitos = RequisitoPazSalvo.objects.filter(mostrar=True, status=True)
                    if not pazsalvo.estado_requisito == 1:
                        for r in requisitos:
                            doc = r.documento_pazsalvo(pazsalvo)
                            if not doc:
                                doc = DocumentoPazSalvo(pazsalvo=pazsalvo, requisito=r, obligatorio = not r.opcional)
                                doc.save(request)
                                log(u'Adiciono documento: %s' % persona, request, "add")
                            else:
                                doc.opcional = not r.opcional
                                doc.save(request)
                    template = get_template("th_hojavida/informacionlaboral/modal/formrequisitos.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'observaciones':
                try:
                    data['idp'] = id = encrypt_id(request.GET['id'])
                    data['pazsalvo'] = PazSalvo.objects.get(id=id)
                    template = get_template('th_hojavida/modal/observacionespazsalvo.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')


            # SANCIONES
            elif action == 'sanciones':
                try:
                    data['title'] = u'Sanciones disciplinarias'
                    url_vars, filtro = f'&action={action}', Q(status=True, personasancion__estado__in=[1,2,3,4],
                                                              personasancion__persona=persona,
                                                              estado__gte=3)
                    listado = IncidenciaSancion.objects.filter(filtro).order_by('-id').distinct()
                    paginator = MiPaginador(listado, 20)
                    page = int(request.GET.get('page', 1))
                    paginator.rangos_paginado(page)
                    data['paging'] = paging = paginator.get_page(page)
                    data['listado'] = paging.object_list
                    data['url_vars'] = url_vars
                    data['estados'] = ESTADO_INCIDENCIA[4:]
                    request.session['viewactivoth'] = ['informacionlaboral', action]
                    return render(request, "th_hojavida/informacionlaboral/sanciones.html", data)
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'respuestadescargo':
                try:
                    per_sancion = PersonaSancion.objects.get(id=encrypt_id(request.GET['id']))
                    data['id']=per_sancion.id
                    data['per_sancion'] = per_sancion
                    template = get_template("th_hojavida/informacionlaboral/modal/formrespuestadescargo.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({'result': False, 'mensaje': f'Ocurrio un error: {ex.__str__()}'})

            elif action == 'editrespuestadescargo':
                try:
                    respuesta = RespuestaDescargo.objects.get(id=encrypt_id(request.GET['id']))
                    data['form'] = RespuestaDescargoForm(instancia=respuesta, initial=model_to_dict(respuesta))
                    data['id'] = respuesta.id
                    template = get_template("ajaxformmodal.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({'result': False, 'mensaje': f'Ocurrio un error: {ex.__str__()}'})


            elif action == 'revisarincidencia':
                try:
                    template, data = get_revisar_audiencia(data, request)
                    return render(request, template, data)
                except Exception as ex:
                    messages.error(request, f'Error: {ex}')

            elif action == 'detalleaudiencia':
                try:
                    audiencia = AudienciaSancion.objects.get(id=encrypt_id(request.GET['id']))
                    data['audiencia'] = audiencia
                    template = get_template('adm_sanciones/modal/involucrados_audiencia.html')
                    return JsonResponse({'result': True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({'result': False, 'mensaje': f'Error: {ex}'})

            elif action == 'firmardocumento':
                try:
                    data['id'] = encrypt_id(request.GET['id'])
                    template = get_template("formfirmaelectronicamasiva.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, 'mensaje': f'Error: {ex}'})

            elif action == 'justificarnoasistencia':
                try:
                    id = encrypt_id(request.GET['id'])
                    idex = encrypt_id(request.GET['idex'])
                    per_aud = PersonaAudienciaSancion.objects.get(id=id)
                    if not per_aud:
                        return JsonResponse({'result': False, 'mensaje': 'Usuario no encontrado en la audiencia'})
                    form = JustificacionPersonaAudienciaForm()
                    if idex == 0:  # editar
                        form.fields['justificacion'].initial = per_aud.justificacion
                        form.fields['archivo'].initial = per_aud.archivo
                    data['id'] = per_aud.id
                    data['form'] = form
                    template = get_template("ajaxformmodal.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return HttpResponseRedirect(request.path)

            elif action == 'detallenoasistencia':
                try:
                    id = encrypt_id(request.GET['id'])
                    per_aud = PersonaAudienciaSancion.objects.get(id=id)
                    data['per_aud'] = per_aud
                    template = get_template('adm_sanciones/modal/detallenoasistencia.html')
                    return JsonResponse({'result': True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({'result': False, 'mensaje': f'Error: {ex}'})

            elif action == 'nofirmaraccionpers':
                try:
                    data['id'] = encrypt_id(request.GET['id'])
                    form = MotivoNoFirmaAccionPersonalForm()
                    data['form'] = form
                    template = get_template("ajaxformmodal.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return HttpResponseRedirect(request.path)


            # OCULTAR POPUP DE NOTIFICACION
            elif action == 'notifyverincidencia':
                try:
                    id = encrypt_id(request.GET['id'])
                    persona_sancion = PersonaSancion.objects.get(id=id)
                    persona_sancion.notificacion = 3
                    persona_sancion.save()
                    request.session['persona_sancion_notificar'] = None
                    url = f'/th_hojavida?action=revisarincidencia&id={encrypt(persona_sancion.incidencia.id)}'
                    url_completa = dominio_sistema_base(request) + url
                    return JsonResponse({'result': True, 'url': f'{url_completa}'})

                except Exception as ex:
                    return JsonResponse({'result': False, 'mensaje': f'Error: {ex}'})

        else:
            try:
                data['title'] = u'Datos personales'
                data['persona_'] = persona = Persona.objects.get(id=persona.id)
                data['perfil'] = persona.mi_perfil()
                data['migrante'] = MigrantePersona.objects.filter(persona=persona).first()
                data['documentopersonal'] = persona.documentos_personales()
                data['reporte_1'] = obtener_reporte('hoja_vida_sagest')
                data['reporte_2'] = obtener_reporte('certificado_ficha_estudiantil')
                request.session['viewactivoth'] = ['informacionpersonal', 'datospersonales']
                return render(request, "th_hojavida/informacionpersonal/datospersonales.html", data)
            except Exception as ex:
                messages.error(request, f'{ex}')


def validar_choque_horario_actividad_gestion(persona, fi, hi, hf, periodo=None):
    listaturnosocupados, choqueturno = [], []
    dia = fi.isoweekday()
    if Profesor.objects.filter(persona=persona, status=True, activo=True).values('id').exists():
        # HORARIO DE DOCENCIA
        profesor = Profesor.objects.filter(persona=persona, status=True, activo=True).order_by('-id').first()
        fActiv = Q(status=True, activo=True, inicio__lte=fi, fin__gte=fi, turno__status=True, dia=dia,
                   detalledistributivo__distributivo__profesor=profesor)
        fClass = Q(materia__visiblehorario=True, materia__fechafinasistencias__gte=fi, status=True, activo=True,
                   inicio__lte=fi, fin__gte=fi, turno__mostrar=True, turno__status=True, dia=dia,
                   materia__profesormateria__profesor_id=profesor.id)
        if periodo:
            if not periodo.es_posgrado() and not periodo.es_admision():
                fClass &= Q(materia__nivel__periodo=periodo)
                fActiv &= Q(detalledistributivo__distributivo__periodo=periodo)
        profesormateria = ProfesorMateria.objects.values_list('materia_id', flat=True).filter(materia__inicio__lte=fi,
                                                                                              materia__fin__gte=fi,
                                                                                              profesor=profesor,
                                                                                              profesor__status=True,
                                                                                              activo=True,
                                                                                              principal=True,
                                                                                              status=True,
                                                                                              materia__status=True,
                                                                                              materia__cerrado=False).exclude(
            materia__nivel__periodo__tipo__id__in=[3, 4]).distinct()
        for m in Materia.objects.filter(pk__in=profesormateria):
            clases = m.clase_set.filter(fClass).order_by('turno__comienza')
            for clase in clases:
                listaturnosocupados.append([clase.turno, m])
                if clases.filter(
                        Q(id=clase.id) & (Q(turno__comienza__range=(hi, hf)) | Q(turno__termina__range=(hi, hf)))):
                    choqueturno.append(clase.turno.id)
        # HORARIO DE ACTIVIDADES
        actividades = ClaseActividad.objects.filter(fActiv).exclude(
            detalledistributivo__criteriodocenciaperiodo__criterio__procesotutoriaacademica=True).distinct().order_by(
            'turno__comienza')
        ids_choque = actividades.values_list('id', flat=True).filter(
            Q(turno__comienza__range=(hi, hf)) | Q(turno__termina__range=(hi, hf)))
        for ac in actividades:
            actividadnombre = ''
            if ac.detalledistributivo.criteriodocenciaperiodo:
                actividadnombre = ac.detalledistributivo.criteriodocenciaperiodo.criterio.nombre
            if ac.detalledistributivo.criterioinvestigacionperiodo:
                actividadnombre = ac.detalledistributivo.criterioinvestigacionperiodo.criterio.nombre
            if ac.detalledistributivo.criteriogestionperiodo:
                actividadnombre = ac.detalledistributivo.criteriogestionperiodo.criterio.nombre
            if ac.id in ids_choque:
                choqueturno.append(ac.turno.id)

            listaturnosocupados.append([ac.turno, actividadnombre])

    return [choqueturno, listaturnosocupados]

def validar_dia_maximo_bitacora(date_,dias):
    """
    Funcion para limitar con el valor recibido del campo días, como día maximo para subir una actividad en la bitacora
    Ejemplo: 01/08/2023 podrá subir hasta el 03/08/2023
    :param datetime
    :return: boolean
    """
    f_today = datetime.now()
    last_day = calendar.monthrange(f_today.year,f_today.month)[1]
    if not date_.month == f_today.month:
        last_date_month = calendar.monthrange(date_.year, date_.month)[1]
        f_extra = date_ + timedelta(days=dias)
        if f_extra.month == f_today.month:
            f_extra = datetime.strptime(f"{last_date_month}/{date_.month}/{date_.year}",'%d/%m/%Y')
        return False, f_extra
    f_extra = date_ + timedelta(days=dias)
    if not f_extra.month == f_today.month:
        f_extra = date_ + timedelta(days=last_day-date_.day)
    # day_count = abs(f_extra.day - f_today.day)
    if date_.date()>=f_today.date():
        return True,date_
    if f_extra.date() < f_today.date():
        return False,f_extra
    else:
        return True,f_extra

def obtener_posicion_y(urlpdf, palabras):
    pdf = SITE_STORAGE + urlpdf
    documento = fitz.open(pdf)
    numpaginafirma = int(documento.page_count) - 1
    with fitz.open(pdf) as document:
        for page_number, page in enumerate(document):
            if page_number == numpaginafirma:
                posicion = page.search_for(palabras)
    valor = posicion[0][1] if posicion else None
    y = None
    if valor:
        y = 5000 - int(valor) - 4150
    return y, numpaginafirma

def generar_carta_decimos(request, decimo):
    data = {}
    data['persona'] = persona = decimo.persona
    data['decimo'] = decimo
    data['fechaactual'] = datetime.now()
    directory_p = os.path.join(MEDIA_ROOT, 'decimos')
    try:
        os.stat(directory_p)
    except:
        os.mkdir(directory_p)
    directory = os.path.join(MEDIA_ROOT, 'decimos', 'carta_decimos')
    try:
        os.stat(directory)
    except:
        os.mkdir(directory)

    nombre_archivo = generar_nombre(f'carta_decimos_{persona.usuario.username}_{decimo.id}', 'generado') + '.pdf'
    valido = conviert_html_to_pdfsave_generic_lotes(
                    request,
                    'th_hojavida/pdf/acta_decimo.html',
                    {
                        'pagesize': 'A4 landscape',
                        'data': data,
                    }, directory, nombre_archivo
                )
    if not valido:
        raise NameError('Error al generar el informe')
    url_archivo = f'/media/decimos/carta_decimos/{nombre_archivo}'
    pdf_path = SITE_STORAGE + url_archivo
    with open(pdf_path.replace('\\', '/'), "rb") as file:
        pdf_content = file.read()
        pdf = io.BytesIO(pdf_content)
    return pdf, url_archivo


def migrar_par_revisor(parrevisor, persona, periodo, request=None):
    try:
        from sga.models import EvidenciaActividadAudi
        from sga.pro_cronograma import CRITERIO_PAR_EVALUADOR
        from inno.models import SubactividadDetalleDistributivo, MigracionEvidenciaActividad
        if object := SubactividadDetalleDistributivo.objects.filter(subactividaddocenteperiodo__criterio=CRITERIO_PAR_EVALUADOR, actividaddetalledistributivo__criterio__distributivo__profesor=persona.profesor(), actividaddetalledistributivo__criterio__distributivo__periodo=periodo, status=True).first():
            evidencia = None
            fr = parrevisor.fecharevision
            criterio = object.actividaddetalledistributivo.criterio
            desde, hasta = date(fr.year, fr.month, 1), date(fr.year, fr.month, calendar.monthrange(fr.year, fr.month)[1])
            if _migracion := parrevisor.migracionevidenciaactividad_set.filter(status=True).first():
                evidencia = _migracion.evidencia
                evidencia.actividad = parrevisor.titulo
                evidencia.archivo = parrevisor.archivo
                evidencia.desde = desde
                evidencia.hasta = hasta
                if not request:
                    evidencia.save()
                else:
                    evidencia.save(request)
            else:
                evidencia = EvidenciaActividadDetalleDistributivo(criterio=criterio, subactividad=object, desde=desde, hasta=hasta, actividad=parrevisor.titulo, archivo=parrevisor.archivo)
                migracion = MigracionEvidenciaActividad(evidencia=evidencia, parrevisor=parrevisor)
                if not request:
                    evidencia.save()
                    migracion.save()
                else:
                    evidencia.save(request)
                    migracion.save(request)
            evidencia and EvidenciaActividadAudi(evidencia=evidencia, archivo=evidencia.archivo).save()
        return True
    except Exception as ex:
        return False
