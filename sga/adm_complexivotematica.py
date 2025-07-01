# -*- coding: UTF-8 -*-
import json
import os
import sys
from datetime import datetime
import random
import io

import openpyxl
import xlsxwriter
import xlwt
from django.contrib.auth.decorators import login_required
from django.core.exceptions import ObjectDoesNotExist
from django.db import transaction, connection, connections
from django.db.models import Q, Avg
from django.forms import model_to_dict
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.shortcuts import render
from django.template.context import Context
from django.template.loader import get_template
from xlwt import *
from django.db.models.aggregates import Sum, Count
from decorators import secure_module, last_access
from inno.forms import FirmaGrupoTitulacionForm,DisertacionExamenComplexivoForm
from inno.models import RequisitoMateriaUnidadIntegracionCurricular, RequisitoIngresoUnidadIntegracionCurricular, \
    GrupoTitulacionIC, MateriaGrupoTitulacion, MateriaTitulacionNota, FirmaGrupoTitulacion, GrupoFirma, \
    MateriaTitulacionFirma, DisertacionFechaPlanificacion, DisertacionTurnoPlanificacion, \
    DisertacionTribunalPlanificacion, DisertacionGrupoPlanificacion, DisertacionMateriaAsignadaPlanificacion, \
    DisertacionAulaPlanificacion
from settings import NOTA_ESTADO_APROBADO, ESTADO_GESTACION, MEDIA_ROOT
from sga.commonviews import adduserdata
from sga.forms import ComplexivoTematicaForm, ComplexivoTribunalCalificador, ComplexivoTematicaObservacionForm, \
    FirmasPeriodoGrupoTitulacionForm, ParticipanteTematicaForm, ConfiguracionComplexivoHabilitaPropuestaForm, InscripcionTitulacionForm
from sga.funciones import log, MiPaginador, puede_realizar_accion, null_to_decimal, null_to_numeric, convertir_fecha, llenar_requisitostitulacion
from sga.funciones_templatepdf import actatribunalcalificacion, rubricatribunalcalificacion, actatitulacioncomplexivo, listadovalidarequisitos
from sga.models import Tematica, ComplexivoTematica, Carrera, ComplexivoGrupoTematica, MatriculaTitulacion, \
    Coordinacion, MESES_CHOICES, ComplexivoDetalleGrupo, PeriodoGrupoTitulacion, Profesor, CARGOS_JURADO_SUSTENTACION, \
    ComplexivoPropuestaPractica, ActividadDetalleDistributivo, ParticipanteGrupoInvestigacion, ParticipanteTematica, \
    Graduado, LineasTematica, \
    AlternativaTitulacion, Inscripcion, TIPO_CELULAR, PracticasPreprofesionalesInscripcion, ModuloMalla, \
    ParticipantesMatrices, AsignaturaMalla, RecordAcademico, CalificacionRubricaTitulacion, \
    CalificacionDetalleRubricaTitulacion, RubricaTitulacion, ComplexivoTematicaGrupoCupo, \
    ActividadDetalleDistributivoCarrera, DetalleDistributivo, \
    ProfesorDistributivoHoras, FirmaPeriodoGrupoTitulacion, Sede, ComplexivoPropuestaPracticaArchivo, \
    ComplexivoAcompanamiento, RubricaTitulacionCabPonderacion, ModeloRubricaTitulacion, \
    CalificacionDetalleModeloRubricaTitulacion, ExamenComlexivoGraduados, ConfiguracionComplexivoHabilitaPropuesta, \
    TurnoTitulacion, Persona, Materia, MateriaAsignada, RequisitoTitulacionMalla, MateriaTitulacion, Malla, \
    DetalleModeloEvaluativo, Aula, Matricula, LaboratorioVirtual
from sga.templatetags.sga_extras import encrypt, pertenecepredecesoratitulacion, predecesoratitulacion
from sga.funcionesxhtml2pdf import conviert_html_to_pdf
from settings import BASE_DIR

@login_required(redirect_field_name='ret', login_url='/loginsga')
@transaction.atomic()
@secure_module
@last_access
def view(request):
    data = {}
    adduserdata(request, data)
    persona = request.session['persona']
    perfilprincipal = request.session['perfilprincipal']
    profesor = perfilprincipal.profesor
    periodo = request.session['periodo']
    miscarreras = Carrera.objects.filter(coordinadorcarrera__in=persona.gruposcarrera(periodo)).distinct()

    # bordes
    borders = Borders()
    borders.left = 1
    borders.right = 1
    borders.top = 1
    borders.bottom = 1
    # estilos para los reportes
    title = easyxf('font: name Arial, bold on , height 200; alignment: horiz centre')
    subtitle = easyxf('font: name Arial, bold on , height 150; alignment: horiz centre')
    normaliz = easyxf('font: name Arial , height 150; align: wrap on,horiz left ')
    normalizsinborde = easyxf('font: name Arial , height 150; align: wrap on,horiz left ')
    normal = easyxf('font: name Arial, height 150; align: wrap on,horiz center ')
    normalcenter = easyxf('font: name Arial, height 150; align: wrap on,vert centre, horiz left ')
    normalsinborde = easyxf('font: name Arial , height 150; align: wrap on,horiz center ')
    stylebnotas = easyxf('font: name Arial, bold on , height 150; alignment: horiz centre')
    stylebnombre = easyxf('font: name Arial, bold on , height 150; align: wrap on, horiz left')
    stylebnotas.borders = borders
    normaliz.borders = borders
    normalcenter.borders = borders
    if request.method == 'POST':
        if 'action' in request.POST:
            action = request.POST['action']

            if action == 'deleteacta':
                try:
                    with transaction.atomic():
                        instancia = MateriaTitulacion.objects.get(pk=int(request.POST['id']))
                        instancia.numeroacta = 0
                        instancia.actacerrada = False
                        instancia.fechaacta = None
                        instancia.numeromemo = None
                        instancia.numeromemogradua = None
                        instancia.actatitulacionfirmada = False
                        instancia.fechatitulacionfirmada = None
                        instancia.archivotitulaciongenerada = None
                        instancia.grupofirma = None
                        instancia.save(request)
                        if MateriaTitulacionFirma.objects.values('id').filter(materiatitulacion=instancia,status=True).exists():
                            mattitulacionfirma = MateriaTitulacionFirma.objects.filter(materiatitulacion=instancia,status=True)[0]
                            mattitulacionfirma.delete()
                        log(u'Elimino acta de titulacion: %s' % instancia, request, "delete")
                        res_json = {"error": False}
                except Exception as ex:
                    transaction.set_rollback(True)
                    res_json = {'error': True, "message": "Error: {}".format(ex)}
                return JsonResponse(res_json, safe=False)

            if action == 'asignacupogrupo':
                try:
                    tematica = ComplexivoTematica.objects.get(pk=int(request.POST['id']))
                    lista = json.loads(request.POST['grupos'])
                    for l in lista:
                        if not ComplexivoTematicaGrupoCupo.objects.filter(pk=int(l['idgrupo']), enuso=True).exists():
                            ComplexivoTematicaGrupoCupo.objects.filter(pk=int(l['idgrupo'])).update(cupoasignado=int(l['valor']))

                    tematica.cupo = tematica.total_cupos_grupos()
                    tematica.save(request)

                    log(u"Actualizó cupos de grupos de la tematica: %s" % tematica, request, "update")
                    return JsonResponse({'result': 'ok'})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': 'bad', 'mensaje': u"Error al guardar los datos"})

            elif action == 'addtematicacupo':
                try:
                    f = ComplexivoTematicaForm(request.POST)
                    if f.is_valid():
                        tematica = ComplexivoTematica.objects.get(pk=request.POST['id'])
                        tematica.cupo = f.cleaned_data['cupo']
                        tematica.tutor = f.cleaned_data['tutor']
                        tematica.maxintegrantes = f.cleaned_data['maxintegrantes']
                        tematica.save(request)
                        log(u"Adiciono tematica: %s" % tematica, request, "add")
                        return JsonResponse({'result': 'ok'})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': 'bad', 'mensaje': u"Error al guardar los datos"})

            elif action == 'updatecupo':
                try:
                    puede_realizar_accion(request, 'sga.puede_editar_cupotematica')
                    tematica = ComplexivoTematica.objects.get(pk=request.POST['id'])
                    if tematica.cantidad_inscritos() <= int(request.POST['vc']):
                        tematica.cupo = int(request.POST['vc'])
                        tematica.save(request)
                        log(u"Adiciono cupo a la línea de investigación: %s" % tematica, request, "update")
                        return JsonResponse({'result': 'ok'})
                    else:
                        return JsonResponse({'result': 'bad', 'mensaje': u"EL cupo no puede ser menor a la cantidad de incritos"})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': 'bad', 'mensaje': u"Error al guardar los datos"})

            elif action == 'actualizacantidadgrupo':
                try:
                    puede_realizar_accion(request, 'sga.puede_editar_cupotematica')

                    if not request.POST['cg'].isdigit():
                        return JsonResponse({'result': 'bad', 'mensaje': u"El valor debe ser numérico"})

                    if int(request.POST['cg']) == 0:
                        return JsonResponse({'result': 'bad', 'mensaje': u"El valor debe ser mayor a cero."})

                    tematica = ComplexivoTematica.objects.get(pk=request.POST['id'])
                    cantidadgrupoorig = tematica.cantidadgrupo

                    if tematica.cantidad_grupo_ocupado() <= int(request.POST['cg']):
                        tematica.cantidadgrupo = int(request.POST['cg'])
                        tematica.save(request)

                        if not tematica.complexivotematicagrupocupo_set.filter(status=True).exists():
                            for n in range(1, tematica.cantidadgrupo + 1):
                                grupocupo = ComplexivoTematicaGrupoCupo(tematica=tematica,
                                                                        numerogrupo=n,
                                                                        cupoasignado=2,
                                                                        enuso=False)
                                grupocupo.save(request)
                        else:
                            if int(request.POST['cg']) > cantidadgrupoorig:
                                dif = int(request.POST['cg']) - cantidadgrupoorig
                                for n in range(cantidadgrupoorig + 1, cantidadgrupoorig + dif + 1):
                                    grupocupo = ComplexivoTematicaGrupoCupo(tematica=tematica,
                                                                            numerogrupo=n,
                                                                            cupoasignado=2,
                                                                            enuso=False)
                                    grupocupo.save(request)
                            else:
                                dif = cantidadgrupoorig - int(request.POST['cg'])
                                grupos = tematica.complexivotematicagrupocupo_set.filter(status=True, enuso=False).order_by('-id')[:dif]
                                for g in grupos:
                                    g.status = False
                                    g.save(request)

                        tematica.cupo = tematica.total_cupos_grupos()
                        tematica.save(request)

                        log(u"Adiciono cantidad de grupos a la línea de investigación: %s" % tematica, request, "update")
                        return JsonResponse({'result': 'ok'})
                    else:
                        return JsonResponse({'result': 'bad', 'mensaje': u"El valor no puede ser menor a la cantidad de grupos en uso"})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': 'bad', 'mensaje': u"Error al guardar los datos"})

            elif action == 'deletegrupo':
                try:
                    grupo = ComplexivoGrupoTematica.objects.get(pk=request.POST['id'])
                    gc = grupo.grupocupo
                    grupo.status = False
                    grupo.activo = False
                    grupo.grupocupo = None
                    grupo.save(request)

                    if gc:
                        if grupo.tematica.complexivotematicagrupocupo_set.filter(id=gc.id):
                            grupocupo = grupo.tematica.complexivotematicagrupocupo_set.filter(id=gc.id)[0]
                            grupocupo.enuso = False
                            grupocupo.save(request)

                    log(u"Elimino Grupo de Estudiantes : %s" % grupo, request, "delete")
                    for detalle in grupo.complexivodetallegrupo_set.all():
                        detalle.status = False
                        detalle.estado = 3
                        detalle.save(request)
                        # detalle.matricula.estado=8
                        # detalle.matricula.status=False
                        # detalle.matricula.save()

                        log(u"Elimino integrante de la línea de investigación : %s" % grupo, request, "delete")
                    return JsonResponse({'result': 'ok'})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': 'bad', 'mensaje': u'Ocurrio un problema al eliminar grupo'})

            elif action == 'deletetribunal':
                try:
                    grupo = ComplexivoGrupoTematica.objects.get(pk=request.POST['id'])
                    grupo.presidentepropuesta = None
                    grupo.secretariopropuesta = None
                    grupo.delegadopropuesta = None
                    grupo.fechadefensa = None
                    grupo.horadefensa = None
                    grupo.lugardefensa = ''
                    grupo.save(request)
                    return JsonResponse({'result': 'ok'})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': 'bad', 'mensaje': u'Ocurrio un problema al eliminar tribunal'})

            elif action == 'reiniciaestado':
                try:
                    item = ComplexivoPropuestaPractica.objects.get(pk=request.POST['id'])
                    item.estado = 1
                    item.save(request)
                    log(u"Reinició estado a detalle %s de archivos en titulación" % item.id, request, "edit")
                    return JsonResponse({'result': 'ok'})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': 'bad', 'mensaje': u'Ocurrio un problema al reiniciar estado'})

            elif action == 'asignartribunal':
                try:
                    f = ComplexivoTribunalCalificador(request.POST)
                    if f.is_valid():
                        grupo = ComplexivoGrupoTematica.objects.get(pk=int(request.POST['id']))
                        if int(f.cleaned_data['presidente']) == 0 or int(f.cleaned_data['secretario']) == 0 or int(f.cleaned_data['delegado']) == 0 or int(f.cleaned_data['moderador']) == 0:
                            return JsonResponse({'result': 'bad', 'mensaje': u'Todos los campos son obligatorios'})
                        if ComplexivoGrupoTematica.objects.filter(moderador_id=f.cleaned_data['moderador'], fechadefensa=f.cleaned_data['fecha'], horadefensa=f.cleaned_data['hora'], status=True).exclude(pk=grupo.id):
                            return JsonResponse({'result': 'bad', 'mensaje': u'Moderador ya se encuentra en otro grupo en la misma fecha y hora'})
                        if grupo.complexivodetallegrupo_set.filter(status=True):
                            for detalle in grupo.complexivodetallegrupo_set.filter(status=True):
                                if detalle.calificacionrubricatitulacion_set.filter(status=True):
                                    for rubricadetalle in detalle.calificacionrubricatitulacion_set.filter(status=True):
                                        if rubricadetalle.tipojuradocalificador == 1:
                                            rubricadetalle.juradocalificador_id = f.cleaned_data['presidente']
                                        if rubricadetalle.tipojuradocalificador == 2:
                                            rubricadetalle.juradocalificador_id = f.cleaned_data['secretario']
                                        if rubricadetalle.tipojuradocalificador == 3:
                                            rubricadetalle.juradocalificador_id = f.cleaned_data['delegado']
                                        rubricadetalle.save(request)

                        grupo.presidentepropuesta_id = f.cleaned_data['presidente']
                        grupo.secretariopropuesta_id = f.cleaned_data['secretario']
                        grupo.delegadopropuesta_id = f.cleaned_data['delegado']
                        grupo.moderador_id = f.cleaned_data['moderador']
                        grupo.fechadefensa = f.cleaned_data['fecha']
                        grupo.lugardefensa = f.cleaned_data['lugar'].upper()
                        grupo.horadefensa = f.cleaned_data['hora']
                        grupo.save(request)
                        lista_items1 = json.loads(request.POST['lista_items1'])
                        if lista_items1:
                            for ltur in lista_items1:
                                grupo.turnotitulacion_id = ltur
                                grupo.save(request)
                        log(u"Adiciono Tribunal al grupo[%s]" % grupo.id, request, "add")
                        return JsonResponse({'result': 'ok'})
                    else:
                        raise NameError('Error')
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': 'bad', 'mensaje': u"Error al guardar los datos"})

            elif action == 'eliminarparticipante':
                try:
                    f = ComplexivoTematicaObservacionForm(request.POST)
                    if f.is_valid():
                        detalle = ComplexivoDetalleGrupo.objects.get(pk=int(request.POST['id']))
                        participante = detalle.matricula
                        if request.POST['tipo'] == 'eliminar':
                            participante.estado = 9
                            if Graduado.objects.filter(inscripcion=participante.inscripcion):
                                graduado = Graduado.objects.get(inscripcion=participante.inscripcion)
                                graduado.delete()
                            detalle.estado = 3
                        else:
                            participante.estado = 1
                            detalle.estado = 1
                        if participante.motivo:
                            participante.motivo = participante.motivo + ' ' + request.POST['observacion']
                        else:
                            participante.motivo = request.POST['observacion']
                        participante.save()
                        if detalle.observacion:
                            detalle.observacion = detalle.observacion + ' ' + request.POST['observacion']
                        else:
                            detalle.observacion = request.POST['observacion']
                        detalle.save(request)
                        log(u"Modifico participante %s, %s" % (participante, request.POST['tipo']), request, "edit")
                        return JsonResponse({'result': 'ok'})
                    else:
                        raise NameError('Error')
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': 'bad', 'mensaje': u"Error al guardar los datos"})

            elif action == 'addparticipante':
                try:
                    lista = json.loads(request.POST['lista'])
                    lista = map(int, lista)
                    participantes = MatriculaTitulacion.objects.filter(pk__in=lista)
                    for participante in participantes:
                        if not ComplexivoDetalleGrupo.objects.filter(status=True, matricula=participante, grupo__activo=True).exists():
                            detalle = ComplexivoDetalleGrupo()
                            detalle.grupo = ComplexivoGrupoTematica.objects.get(pk=request.POST['id'])
                            detalle.matricula = participante
                            detalle.fechainscripcion = datetime.now()
                            detalle.save(request)
                            log(u"Añadio participante : %s" % detalle, request, "add")
                    return JsonResponse({'result': 'ok'})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': 'bad', 'mensaje': u'Ocurrio un problema al añadir integrante'})

            elif action == 'addgrupo':
                try:
                    lista = json.loads(request.POST['lista'])
                    lista = map(int, lista)

                    participantes = MatriculaTitulacion.objects.filter(pk__in=lista)
                    grupo = ComplexivoGrupoTematica()
                    grupo.tematica_id = request.POST['id']
                    grupo.save(request)
                    log(u"Creo grupo para desarrollar la línea de investigación : %s" % grupo, request, "add")
                    for participante in participantes:
                        if not ComplexivoDetalleGrupo.objects.filter(status=True, matricula=participante, grupo__activo=True).exists():
                            detalle = ComplexivoDetalleGrupo()
                            detalle.grupo = grupo
                            detalle.matricula = participante
                            detalle.fechainscripcion = datetime.now()
                            detalle.save(request)
                            log(u"Añadio integrante a grupo : %s" % detalle, request, "add")
                    return JsonResponse({'result': 'ok'})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': 'bad', 'mensaje': u'Ocurrio un problema al añadir integrante'})

            elif action == 'deleteintegrante':
                try:
                    grupo = ComplexivoGrupoTematica.objects.get(pk=request.POST['id'])
                    lista = json.loads(request.POST['lista'])
                    lista = map(int, lista)
                    ComplexivoDetalleGrupo.objects.filter(pk__in=lista).update(status=False, estado=3)
                    log(u"Elimino integrante de grupo : %s" % grupo, request, "delete")
                    if not grupo.tiene_participantes():
                        gc = grupo.grupocupo
                        grupo.status = False
                        grupo.grupocupo = None
                        grupo.save(request)

                        if gc:
                            grupocupo = grupo.tematica.complexivotematicagrupocupo_set.filter(id=gc.id)[0]
                            grupocupo.enuso = False
                            grupocupo.save(request)

                        log(u"Se elimino el grupo por quedarse sin integrantes : %s" % grupo, request, "delete")
                    return JsonResponse({'result': 'ok'})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': 'bad', 'mensaje': u'Ocurrio un problema al eliminar participante'})

            elif action == 'cambiaestadosubirarchivofinal':
                try:
                    complexivogrupotematica = ComplexivoGrupoTematica.objects.get(pk=int(encrypt(request.POST['id_grupo'])))
                    if complexivogrupotematica.subirarchivofinalgrupo:
                        complexivogrupotematica.subirarchivofinalgrupo = False
                    else:
                        complexivogrupotematica.subirarchivofinalgrupo = True
                        complexivogrupotematica.estadoarchivofinalgrupo = 1
                    complexivogrupotematica.save(request)
                    return JsonResponse({'result': 'ok', 'valor': complexivogrupotematica.subirarchivofinalgrupo})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

            elif action == 'cambiaestadopropuesta':
                try:
                    complexivogrupotematica = ComplexivoGrupoTematica.objects.get(pk=int(encrypt(request.POST['id_grupo'])))
                    if complexivogrupotematica.subirpropuesta:
                        complexivogrupotematica.subirpropuesta = False
                        complexivogrupotematica.fechasubirpropuesta = None

                    else:
                        complexivogrupotematica.subirpropuesta = True
                        complexivogrupotematica.fechasubirpropuesta = datetime.now()
                    complexivogrupotematica.save(request)
                    log(u'Cambió estado subir propuesta a grupo %s' % complexivogrupotematica, request, "edit")

                    return JsonResponse({'result': 'ok', 'valor': complexivogrupotematica.subirarchivofinalgrupo})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

            elif action == 'cambiaestadocerraracta':
                try:
                    integrante = ComplexivoDetalleGrupo.objects.get(pk=int(encrypt(request.POST['idinte'])))
                    if integrante.actacerrada:
                        integrante.actacerrada = False
                        integrante.califico = False
                        integrante.estadotribunal = 1
                        integrante.matricula.estado = 1
                        integrante.matricula.save(request)
                        integrante.save(request)
                    else:
                        integrante.actacerrada = True
                        integrante.califico = True

                        grupo = ComplexivoGrupoTematica.objects.get(pk=integrante.grupo.id, status=True)
                        for detalle in grupo.complexivodetallegrupo_set.filter(Q(status=True), (
                                Q(matricula__estado=1) | Q(matricula__estado=10) | Q(matricula__estado=9))):
                            if detalle.matricula.cumplerequisitos == 2 and detalle.matricula.estado == 10 and detalle.rubrica:
                                if detalle.matricula.inscripcion.completo_malla():
                                    if not detalle.matricula.inscripcion.graduado_set.filter(status=True):
                                        graduado = Graduado(inscripcion=detalle.matricula.inscripcion,
                                                            decano=None,
                                                            notafinal=0,
                                                            nombretitulo='',
                                                            horastitulacion=0,
                                                            creditotitulacion=0,
                                                            creditovinculacion=0,
                                                            creditopracticas=0,
                                                            fechagraduado=None,
                                                            horagraduacion=None,
                                                            fechaactagrado=None,
                                                            profesor=None,
                                                            integrantetribunal=None,
                                                            docentesecretario=None,
                                                            secretariageneral=None,
                                                            representanteestudiantil=None,
                                                            representantedocente=None,
                                                            representantesuplentedocente=None,
                                                            representanteservidores=None,
                                                            matriculatitulacion=None,
                                                            codigomecanismotitulacion=None,
                                                            asistentefacultad=None,
                                                            estadograduado=False,
                                                            docenteevaluador1=None,
                                                            docenteevaluador2=None,
                                                            directorcarrera=None,
                                                            tematesis='')
                                        graduado.save()
                                    if Graduado.objects.filter(Q(status=True), (Q(inscripcion=detalle.matricula.inscripcion) | Q(matriculatitulacion=detalle.matricula))).exists():
                                        graduado = Graduado.objects.get(Q(status=True), Q(inscripcion=detalle.matricula.inscripcion) | Q(matriculatitulacion=detalle.matricula))
                                        notapropuesta = float(detalle.calificacion)
                                        # record = detalle.matricula.inscripcion.promedio_record()
                                        if detalle.matricula.alternativa.tipotitulacion.tipo == 2:
                                            detalleexa = detalle.matricula.complexivoexamendetalle_set.filter(status=True, estado=3)[0]
                                            notaexamen = float(detalleexa.notafinal)
                                            notafinal = null_to_numeric(((notaexamen + notapropuesta) / 2), 2)
                                        elif detalle.matricula.alternativa.tipotitulacion.tipo == 1:
                                            notafinal = null_to_numeric((notapropuesta), 2)
                                        #     nota = null_to_numeric((notapropuesta), 2)
                                        # notafinal = null_to_numeric(((nota + record)/2), 2)
                                        graduado.promediotitulacion = notafinal
                                        graduado.estadograduado = True
                                        graduado.save(request)
                                        log(u'Graduo al estudiante: %s con nota final de: %s, cerro acta de calificación el docente: %s' % (graduado.inscripcion, str(graduado.promediotitulacion), persona), request,
                                            "grad")
                                        if detalle.matricula.alternativa.tipotitulacion.tipo == 2:
                                            if detalle.matricula.complexivoexamendetalle_set.filter(status=True, estado=3).exists():
                                                detalleexamen = detalle.matricula.complexivoexamendetalle_set.filter(status=True, estado=3)[0]
                                                adicionar_nota_complexivo(graduado.id, notapropuesta, detalleexamen.examen.fechaexamen, request)
                                            else:
                                                detalleexamen = detalle.matricula.complexivoexamendetalle_set.filter(status=True)[0]
                                                adicionar_nota_complexivo(graduado.id, notapropuesta, detalleexamen.examen.fechaexamen, request)
                                            detalle.actacerrada = True
                                            detalle.save(request)
                                            grupo.cerrado = True
                                            grupo.save(request)
                                            log(u'Cerro acta de calificación el docente: %s' % (persona), request, "edit")
                                        else:
                                            detalle.actacerrada = True
                                            detalle.save(request)
                                            grupo.cerrado = True
                                            grupo.save(request)
                                else:
                                    detalle.actacerrada = True
                                    detalle.save(request)
                                    grupo.cerrado = True
                                    grupo.save(request)
                            else:
                                detalle.califico = True
                                detalle.actacerrada = True
                                if detalle.matricula.cumplerequisitos == 3:
                                    detalle.matriculaaptahistorico = False
                                detalle.save(request)
                                detalle.matricula.estado = 9
                                detalle.matricula.save(request)
                                grupo.cerrado = True
                                grupo.save(request)

                        integrante.save(request)
                        integrante.actualiza_estado()
                    return JsonResponse({'result': 'ok'})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

            elif action == 'adicionarreplica':
                try:
                    codigogruporeplica = request.POST['codigogruporeplica']
                    codparticipantereplica = request.POST['codparticipantereplica']

                    complexivo = ComplexivoGrupoTematica.objects.get(pk=codigogruporeplica)
                    if complexivo.participantes().count() == 1:
                        complexivo.estado = 2
                        complexivo.save()

                    replicagrupo = complexivo
                    replicagrupo.pk = None
                    replicagrupo.estado = 1
                    replicagrupo.replica = True
                    replicagrupo.presidentepropuesta = None
                    replicagrupo.secretariopropuesta = None
                    replicagrupo.delegadopropuesta = None
                    replicagrupo.fechadefensa = None
                    replicagrupo.horadefensa = None
                    replicagrupo.lugardefensa = ''
                    replicagrupo.enlacevideo = ''
                    replicagrupo.cerrado = False
                    replicagrupo.save(request)

                    complexivotematica = ComplexivoGrupoTematica.objects.get(pk=codigogruporeplica)
                    for acompa in complexivotematica.complexivoacompanamiento_set.filter(status=True):
                        acompanamiento = ComplexivoAcompanamiento(grupo=replicagrupo,
                                                                  fecha=acompa.fecha,
                                                                  horainicio=acompa.horainicio,
                                                                  horafin=acompa.horafin,
                                                                  horas=acompa.horas,
                                                                  observaciones=acompa.observaciones,
                                                                  enlacevideo=acompa.enlacevideo)
                        acompanamiento.save(request)
                    for propuestas in complexivotematica.complexivopropuestapractica_set.filter(status=True):
                        propu = ComplexivoPropuestaPractica(grupo=replicagrupo,
                                                            observacion=propuestas.observacion,
                                                            estado=propuestas.estado,
                                                            porcentajeurkund=propuestas.porcentajeurkund,
                                                            fecharevision=propuestas.fecharevision)
                        propu.save(request)
                        for archi in propuestas.complexivopropuestapracticaarchivo_set.filter(status=True):
                            propuestaarchivo = ComplexivoPropuestaPracticaArchivo(propuesta=propu,
                                                                                  archivo=archi.archivo,
                                                                                  tipo=archi.tipo,
                                                                                  fecha=archi.fecha)
                            propuestaarchivo.save(request)

                    participantegrupo = ComplexivoDetalleGrupo.objects.get(pk=codparticipantereplica)
                    replicaparticipante = participantegrupo
                    replicaparticipante.pk = None
                    replicaparticipante.grupo = replicagrupo
                    replicaparticipante.calpresidente = 0
                    replicaparticipante.calsecretaria = 0
                    replicaparticipante.caldelegado = 0
                    replicaparticipante.calificacion = 0
                    replicaparticipante.califico = False
                    replicaparticipante.actacerrada = False
                    replicaparticipante.numeroacta = 0
                    replicaparticipante.matriculaaptahistorico = True
                    replicaparticipante.actatribunalgenerada = False
                    replicaparticipante.aptogrupo = True
                    replicaparticipante.estadotribunal = 1
                    replicaparticipante.rubrica = None
                    replicaparticipante.save(request)

                    replicaparticipante.matricula.estado = 1
                    replicaparticipante.matricula.save()
                    return JsonResponse({"result": "ok"})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": "Error al guardar los datos."})

            elif action == 'adicionarreplicatodos':
                try:
                    codigogruporeplica = request.POST['codigogruporeplica']
                    complexivo = ComplexivoGrupoTematica.objects.get(pk=codigogruporeplica)
                    complexivo.estado = 2
                    complexivo.save()

                    replicagrupo = complexivo
                    replicagrupo.pk = None
                    replicagrupo.estado = 1
                    replicagrupo.replica = True
                    replicagrupo.presidentepropuesta = None
                    replicagrupo.secretariopropuesta = None
                    replicagrupo.delegadopropuesta = None
                    replicagrupo.fechadefensa = None
                    replicagrupo.horadefensa = None
                    replicagrupo.lugardefensa = ''
                    replicagrupo.enlacevideo = ''
                    replicagrupo.cerrado = False
                    replicagrupo.save(request)

                    complexivotematica = ComplexivoGrupoTematica.objects.get(pk=codigogruporeplica)
                    for acompa in complexivotematica.complexivoacompanamiento_set.filter(status=True):
                        acompanamiento = ComplexivoAcompanamiento(grupo=replicagrupo,
                                                                  fecha=acompa.fecha,
                                                                  horainicio=acompa.horainicio,
                                                                  horafin=acompa.horafin,
                                                                  horas=acompa.horas,
                                                                  observaciones=acompa.observaciones,
                                                                  enlacevideo=acompa.enlacevideo)
                        acompanamiento.save(request)
                    for propuestas in complexivotematica.complexivopropuestapractica_set.filter(status=True):
                        propu = ComplexivoPropuestaPractica(grupo=replicagrupo,
                                                            observacion=propuestas.observacion,
                                                            estado=propuestas.estado,
                                                            porcentajeurkund=propuestas.porcentajeurkund,
                                                            fecharevision=propuestas.fecharevision)
                        propu.save(request)
                        for archi in propuestas.complexivopropuestapracticaarchivo_set.filter(status=True):
                            propuestaarchivo = ComplexivoPropuestaPracticaArchivo(propuesta=propu,
                                                                                  archivo=archi.archivo,
                                                                                  tipo=archi.tipo,
                                                                                  fecha=archi.fecha)
                            propuestaarchivo.save(request)

                    for listaparticipantes in complexivotematica.complexivodetallegrupo_set.filter(status=True):
                        participantegrupo = listaparticipantes
                        replicaparticipante = participantegrupo
                        replicaparticipante.pk = None
                        replicaparticipante.grupo = replicagrupo
                        replicaparticipante.calpresidente = 0
                        replicaparticipante.calsecretaria = 0
                        replicaparticipante.caldelegado = 0
                        replicaparticipante.calificacion = 0
                        replicaparticipante.califico = False
                        replicaparticipante.actacerrada = False
                        replicaparticipante.numeroacta = 0
                        replicaparticipante.rubrica = None
                        replicaparticipante.estadotribunal = 1
                        replicaparticipante.matriculaaptahistorico = True
                        replicaparticipante.actatribunalgenerada = False
                        replicaparticipante.save(request)

                    return JsonResponse({"result": "ok"})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": "Error al guardar los datos."})

            elif action == 'editfechaacta':
                try:
                    fechacta = request.POST['fechacta']
                    codparticipante = request.POST['codparticipante']
                    complexivo = ComplexivoDetalleGrupo.objects.get(pk=codparticipante)
                    complexivo.fechaacta = fechacta
                    complexivo.save(request)
                    return JsonResponse({"result": "ok"})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": "Error al guardar los datos."})

            elif action == 'cambiartutor':
                try:
                    if 'idt' in request.POST and 'idp' in request.POST:
                        tematica = ComplexivoTematica.objects.get(pk=request.POST['idt'])
                        tutor = tematica.tutor
                        participante = ParticipanteTematica.objects.get(id=request.POST['idp'], status=True)
                        tematica.tutor = participante
                        tematica.save(request)
                        log(u"Cambio, el tutor %s de la tematica %s por %s" % (tutor, tematica, participante), request, "edit")
                        return JsonResponse({'result': 'ok'})
                    else:
                        return JsonResponse({'result': 'bad', 'mensaje': u'Error al guardar los datos'})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': 'bad', 'mensaje': u'Ocurrio un problema al eliminar participante'})

            elif action == 'actualizar_graduado':
                try:
                    if 'id' in request.POST:
                        grupo = ComplexivoGrupoTematica.objects.get(pk=request.POST['id'])
                        for integrante in grupo.participantes():
                            if Graduado.objects.filter(status=True, inscripcion=integrante.matricula.inscripcion, matriculatitulacion=integrante.matricula).exists():
                                graduado = Graduado.objects.filter(status=True, inscripcion=integrante.matricula.inscripcion, matriculatitulacion=integrante.matricula)[0]
                                if graduado:
                                    graduado.profesor = grupo.presidentepropuesta
                                    graduado.integrantetribunal = grupo.delegadopropuesta
                                    graduado.docentesecretario = grupo.secretariopropuesta
                                    graduado.save(request)
                        return JsonResponse({'result': 'ok'})
                    else:
                        return JsonResponse({'result': 'bad', 'mensaje': u'Error al guardar los datos'})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': 'bad', 'mensaje': u'Ocurrio un problema al eliminar participante'})

            elif action == 'pdfactatribunalcalificaciones':
                try:
                    data = {}
                    data['fechaactual'] = datetime.now()
                    fechainiciaactagenerar = datetime.strptime('2020-05-10', '%Y-%m-%d').date()
                    data['participante'] = participante = ComplexivoDetalleGrupo.objects.get(pk=request.POST['id'])
                    if not participante.actatribunalgenerada:
                        if participante.grupo.fechadefensa > fechainiciaactagenerar:
                            participante.actatribunalgenerada = True
                            participante.numeroacta = ComplexivoDetalleGrupo.objects.filter(status=True).order_by('-numeroacta')[0].numeroacta + 1
                            participante.fechaacta = datetime.now().date()
                            participante.save(request)
                    data['detallecalificacion'] = detallecalificacion = participante.calificacionrubricatitulacion_set.filter(status=True).order_by('tipojuradocalificador')
                    data['promediopuntajetrabajointegral'] = detallecalificacion.values_list('puntajetrabajointegral').aggregate(promedio=Avg('puntajetrabajointegral'))['promedio']
                    data['promediodefensaoral'] = detallecalificacion.values_list('puntajedefensaoral').aggregate(promedio=Avg('puntajedefensaoral'))['promedio']
                    data['promediofinal'] = promediofinal = null_to_decimal(detallecalificacion.values_list('puntajerubricas').aggregate(promedio=Avg('puntajerubricas'))['promedio'], 2)
                    if participante.matricula.alternativa.tipotitulacion.tipo == 2:
                        data['finalcomplexivo'] = null_to_decimal(((float(participante.notafinal()) + float(promediofinal)) / 2), 2)

                    return conviert_html_to_pdf('adm_complexivotematica/acta_calificaciontribunal_pdf.html',
                                                {
                                                    'pagesize': 'A4',
                                                    'data': data,
                                                }
                                                )
                except Exception as ex:
                    pass

            elif action == 'pdfactatribunalcalificacionesnew':
                try:
                    iddetallegrupo = request.POST['id']
                    actatribunal = actatribunalcalificacion(iddetallegrupo)
                    return actatribunal
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

            elif action == 'pdfrubricacalificaciones':
                try:
                    data = {}
                    lista = []
                    data['fechaactual'] = datetime.now()
                    data['participante'] = participante = ComplexivoDetalleGrupo.objects.get(pk=request.POST['id'])
                    data['calificacionrubricatitulacion'] = calificacionrubricatitulacion = participante.calificacionrubricatitulacion_set.filter(status=True).order_by('tipojuradocalificador')
                    data['numerotribunales'] = calificacionrubricatitulacion.count()
                    data['promediofinal'] = null_to_decimal(calificacionrubricatitulacion.values_list('puntajerubricas').aggregate(promedio=Avg('puntajerubricas'))['promedio'], 2)
                    rubricasevaluadas = RubricaTitulacion.objects.select_related().filter(tipotitulacion=participante.matricula.alternativa.tipotitulacion.tipo, status=True).order_by('id')
                    for rubrica in rubricasevaluadas:
                        puntajepresidente = rubrica.calificaciondetallerubricatitulacion_set.filter(calificacionrubrica__complexivodetallegrupo=participante, status=True).order_by('calificacionrubrica__tipojuradocalificador')
                        lista.append([rubrica, puntajepresidente])
                    data['rubricasevaluadas'] = lista
                    return conviert_html_to_pdf('adm_complexivotematica/pdfrubricacalificaciones_pdf.html',
                                                {
                                                    'pagesize': 'A4',
                                                    'data': data,
                                                }
                                                )
                except Exception as ex:
                    pass

            elif action == 'pdfactatitulacioncomplexivo':
                try:
                    idasignadotitulacion = request.POST['idmateriaasign']
                    asignado = MateriaTitulacion.objects.get(pk=idasignadotitulacion)
                    # if asignado.materiaasignada.materia.grupotitulacionic_set.filter(tiporubrica=4, status=True):
                    #     grupo = asignado.materiaasignada.materia.grupotitulacionic_set.filter(tiporubrica=4, status=True)[0]

                    if asignado.actatitulacionfirmada:
                        qrname = 'qr_actatitulacion_' + str(asignado.id)
                        actatribunal = 'https://sga.unemi.edu.ec//media/qrcode/actatitulacion/' + qrname + '_firmado.pdf'
                        # actatribunal = 'http://127.0.0.1:8000/media/qrcode/actatitulacion/' + qrname + '_firmado.pdf'
                    else:
                        actatribunal = actatitulacioncomplexivo(idasignadotitulacion)
                        if actatribunal:
                            qrname = 'qr_actatitulacion_' + str(asignado.id)
                            actatribunal = 'https://sga.unemi.edu.ec//media/qrcode/actatitulacion/' + qrname + '.pdf'
                            # actatribunal = 'http://127.0.0.1:8000/media/qrcode/actatitulacion/' + qrname + '.pdf'
                    return JsonResponse({"result": "ok", 'url': actatribunal})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

            elif action == 'pdfrubricacalificacionesnew':
                try:
                    iddetallegrupo = request.POST['id']
                    rubricatribunal = rubricatribunalcalificacion(iddetallegrupo)
                    return rubricatribunal
                except Exception as ex:
                    pass

            elif action == 'addlinkgrupo':
                try:
                    grupo = ComplexivoGrupoTematica.objects.get(pk=request.POST['codgrupo'])
                    grupo.enlacevideo = request.POST['observacion']
                    grupo.save(request)
                    return JsonResponse({"result": "ok"})
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})
                    pass

            elif action == 'reportetutoresgrupo_nuevo_distributivo':
                try:
                    data = {}
                    listamatriz = []
                    listamatrizcoordinacion = []
                    data['fechaactual'] = datetime.now()
                    fecha = datetime.today().date()
                    data['fecha'] = str(fecha.day) + " de " + str(MESES_CHOICES[fecha.month - 1][1]).lower() + " del " + str(fecha.year)
                    idper = int(request.POST['idper'])
                    codifacu = int(request.POST['codifacu'])
                    codicarr = int(request.POST['codicarr'])
                    data['periodogrupo'] = periodogrupo = PeriodoGrupoTitulacion.objects.get(pk=request.POST['idper'])
                    data['listadofirmas'] = periodogrupo.firmaperiodogrupotitulacion_set.filter(status=True).order_by('tipofirma')
                    if idper > 0 and codifacu == 0 and codicarr == 0:
                        tematicas = ProfesorDistributivoHoras.objects.values_list('profesor__persona__id', 'profesor__persona__apellido1', 'profesor__persona__apellido2', 'profesor__persona__nombres', 'carrera__id', 'carrera__nombre', 'carrera__coordinacion__id', 'carrera__coordinacion__nombre', 'horasdocencia', 'horasinvestigacion').filter(periodo=periodo, status=True).distinct().order_by('carrera__nombre', 'profesor__persona__apellido1')
                    if codifacu > 0 and codicarr == 0:
                        tematicas = ProfesorDistributivoHoras.objects.values_list('profesor__persona__id', 'profesor__persona__apellido1', 'profesor__persona__apellido2', 'profesor__persona__nombres', 'carrera__id', 'carrera__nombre', 'carrera__coordinacion__id', 'carrera__coordinacion__nombre', 'horasdocencia', 'horasinvestigacion').filter(carrera__coordinacion__id=codifacu, periodo_id=idper, status=True, profesor__status=True).distinct().order_by('carrera__nombre', 'tutor__participante__persona__apellido1')
                    if codicarr > 0:
                        tematicas = ProfesorDistributivoHoras.objects.values_list('profesor__persona__id', 'profesor__persona__apellido1', 'profesor__persona__apellido2', 'profesor__persona__nombres', 'carrera__id', 'carrera__nombre', 'carrera__coordinacion__id', 'carrera__coordinacion__nombre', 'horasdocencia', 'horasinvestigacion').filter(carrera__id=codicarr, periodo_id=idper, status=True, profesor__status=True).distinct().order_by('carrera__nombre', 'tutor__participante__persona__apellido1')

                    totalhorasgrupo = 0
                    nombrecriterio = ""
                    row_num = 1

                    for tematica in tematicas:
                        act_carrera = tematica[5]
                        horasactividad = 0
                        iddistri = 0
                        idactivi = 0
                        detallehoras = 0

                        if DetalleDistributivo.objects.filter(distributivo__profesor__persona__id=tematica[0], distributivo__periodo=periodo, criteriodocenciaperiodo__criterio__id__in=[29, 99], status=True):
                            distri = DetalleDistributivo.objects.get(distributivo__profesor__persona__id=tematica[0], distributivo__periodo=periodo, criteriodocenciaperiodo__criterio__id__in=[29, 99], status=True)
                            horasactividad = distri.horas
                            nombrecriterio = distri.criteriodocenciaperiodo.criterio.nombre
                            iddistri = distri.id

                            horasinve = tematica[8]
                            horasinve2 = tematica[9]

                            if ActividadDetalleDistributivo.objects.filter(criterio=iddistri, status=True).values_list('id', flat=True):
                                activi = ActividadDetalleDistributivo.objects.get(criterio=iddistri, status=True)
                                idactivi = activi.id
                                if ActividadDetalleDistributivoCarrera.objects.filter(actividaddetalle=idactivi, carrera=tematica[4], status=True):
                                    detalledistri = ActividadDetalleDistributivoCarrera.objects.get(actividaddetalle=idactivi, carrera=tematica[4], status=True)
                                    detallehoras = detalledistri.horas
                                    act_carrera = detalledistri.carrera.nombre

                        totalhoras = ComplexivoGrupoTematica.objects.filter(status=True, tematica__carrera_id=tematica[4], tematica__periodo_id=idper, activo=True, tematica__tutor__participante__persona_id=tematica[0])
                        totalhorasgrupo = totalhorasgrupo + totalhoras.count()
                        totalh = int(totalhoras.count())
                        listamatriz.append([tematica[0], tematica[1], tematica[2], tematica[3], tematica[4], act_carrera, 1, tematica[6], tematica[7], int(horasactividad), int(detallehoras), row_num, horasactividad])
                        row_num += 1
                    data['listacoordinaciones'] = Coordinacion.objects.filter(pk__in=tematicas.values_list('carrera__coordinacion__id', flat=True)).distinct()
                    listacarreras = None

                    if codicarr > 0:
                        listacarreras = Carrera.objects.filter(pk__in=tematicas.values_list('carrera__id', flat=True)).distinct()

                    data['nombrecriterio'] = nombrecriterio
                    data['listacarreras'] = listacarreras
                    data['tematicas'] = listamatriz
                    data['sede'] = Sede.objects.get(pk=1)
                    data['periodo'] = periodo
                    data['totalhorasgrupo'] = totalhorasgrupo
                    return conviert_html_to_pdf(
                        'adm_complexivotematica/reportetutoresgrupo_pdf_antiguos.html',
                        {
                            'pagesize': 'A4',
                            'data': data,
                        }
                    )
                except Exception as ex:
                    pass

            elif action == 'reportetutoresgrupo_pdf':
                try:
                    data = {}
                    listamatriz = []
                    listamatrizcoordinacion = []
                    data['fechaactual'] = datetime.now()
                    fecha = datetime.today().date()
                    data['fecha'] = str(fecha.day) + " de " + str(MESES_CHOICES[fecha.month - 1][1]).lower() + " del " + str(fecha.year)
                    idper = int(request.POST['idper'])
                    codifacu = int(request.POST['codifacu'])
                    codicarr = int(request.POST['codicarr'])
                    data['periodogrupo'] = periodogrupo = PeriodoGrupoTitulacion.objects.get(pk=request.POST['idper'])
                    data['listadofirmas'] = periodogrupo.firmaperiodogrupotitulacion_set.filter(status=True).order_by('tipofirma')
                    if idper > 0 and codifacu == 0 and codicarr == 0:
                        tematicas = ComplexivoTematica.objects.values_list('tutor__participante__persona__id', 'tutor__participante__persona__apellido1', 'tutor__participante__persona__apellido2', 'tutor__participante__persona__nombres', 'carrera__id', 'carrera__nombre', 'carrera__coordinacion__id', 'carrera__coordinacion__nombre').filter(periodo_id=idper, status=True, tutor__status=True).distinct().order_by('carrera__nombre', 'tutor__participante__persona__apellido1')
                    if codifacu > 0 and codicarr == 0:
                        tematicas = ComplexivoTematica.objects.values_list('tutor__participante__persona__id', 'tutor__participante__persona__apellido1', 'tutor__participante__persona__apellido2', 'tutor__participante__persona__nombres', 'carrera__id', 'carrera__nombre', 'carrera__coordinacion__id', 'carrera__coordinacion__nombre').filter(carrera__coordinacion__id=codifacu, periodo_id=idper, status=True, tutor__status=True).distinct().order_by('carrera__nombre', 'tutor__participante__persona__apellido1')
                    if codicarr > 0:
                        tematicas = ComplexivoTematica.objects.values_list('tutor__participante__persona__id', 'tutor__participante__persona__apellido1', 'tutor__participante__persona__apellido2', 'tutor__participante__persona__nombres', 'carrera__id', 'carrera__nombre', 'carrera__coordinacion__id', 'carrera__coordinacion__nombre').filter(carrera__id=codicarr, periodo_id=idper, status=True, tutor__status=True).distinct().order_by('carrera__nombre', 'tutor__participante__persona__apellido1')

                    totalhorasgrupo = 0
                    nombrecriterio = ""
                    row_num = 1

                    for tematica in tematicas:
                        horasactividad = 0
                        iddistri = 0
                        idactivi = 0
                        detallehoras = 0

                        if DetalleDistributivo.objects.filter(distributivo__profesor__persona__id=tematica[0], distributivo__periodo=periodo, criteriodocenciaperiodo__criterio__id__in=[29, 99, 125], status=True):
                            distri = DetalleDistributivo.objects.get(distributivo__profesor__persona__id=tematica[0], distributivo__periodo=periodo, criteriodocenciaperiodo__criterio__id__in=[29, 99, 125], status=True)
                            horasactividad = distri.horas
                            nombrecriterio = distri.criteriodocenciaperiodo.criterio.nombre
                            iddistri = distri.id
                            if ActividadDetalleDistributivo.objects.filter(criterio=iddistri, status=True).values_list('id', flat=True):
                                activi = ActividadDetalleDistributivo.objects.get(criterio=iddistri, status=True)
                                idactivi = activi.id
                                if ActividadDetalleDistributivoCarrera.objects.filter(actividaddetalle=idactivi, carrera=tematica[4], status=True):
                                    detalledistri = ActividadDetalleDistributivoCarrera.objects.get(actividaddetalle=idactivi, carrera=tematica[4], status=True)
                                    detallehoras = detalledistri.horas

                        totalhoras = ComplexivoGrupoTematica.objects.filter(status=True, tematica__carrera_id=tematica[4], tematica__periodo_id=idper, activo=True, tematica__tutor__participante__persona_id=tematica[0])
                        totalhorasgrupo = totalhorasgrupo + totalhoras.count()
                        totalh = int(totalhoras.count())
                        listamatriz.append([tematica[0], tematica[1], tematica[2], tematica[3], tematica[4], tematica[5], totalh, tematica[6], tematica[7], int(horasactividad), int(detallehoras), row_num])
                        row_num += 1
                    data['listacoordinaciones'] = Coordinacion.objects.filter(pk__in=tematicas.values_list('carrera__coordinacion__id', flat=True)).distinct()
                    listacarreras = None

                    if codicarr > 0:
                        listacarreras = Carrera.objects.filter(pk__in=tematicas.values_list('carrera__id', flat=True)).distinct()

                    data['nombrecriterio'] = nombrecriterio
                    data['listacarreras'] = listacarreras
                    data['tematicas'] = listamatriz
                    data['sede'] = Sede.objects.get(pk=1)
                    data['periodo'] = periodo
                    data['totalhorasgrupo'] = totalhorasgrupo
                    return conviert_html_to_pdf(
                        'adm_complexivotematica/reportetutoresgrupo_pdf.html',
                        {
                            'pagesize': 'A4',
                            'data': data,
                        }
                    )
                except Exception as ex:
                    pass

            elif action == 'reportetutoresgrupo_pdf_antiguo':
                try:
                    data = {}
                    listamatriz = []
                    listamatrizcoordinacion = []
                    data['fechaactual'] = datetime.now()
                    fecha = datetime.today().date()
                    data['fecha'] = str(fecha.day) + " de " + str(MESES_CHOICES[fecha.month - 1][1]).lower() + " del " + str(fecha.year)
                    idper = int(request.POST['idper'])
                    codifacu = int(request.POST['codifacu'])
                    codicarr = int(request.POST['codicarr'])
                    data['periodogrupo'] = periodogrupo = PeriodoGrupoTitulacion.objects.get(pk=request.POST['idper'])
                    data['listadofirmas'] = periodogrupo.firmaperiodogrupotitulacion_set.filter(status=True).order_by('tipofirma')
                    if idper > 0 and codifacu == 0 and codicarr == 0:
                        tematicas = ComplexivoTematica.objects.values_list('tutor__participante__persona__id', 'tutor__participante__persona__apellido1', 'tutor__participante__persona__apellido2', 'tutor__participante__persona__nombres', 'carrera__id', 'carrera__nombre', 'carrera__coordinacion__id', 'carrera__coordinacion__nombre').filter(periodo_id=idper, status=True, tutor__status=True).distinct().order_by('carrera__nombre', 'tutor__participante__persona__apellido1')
                    if codifacu > 0 and codicarr == 0:
                        tematicas = ComplexivoTematica.objects.values_list('tutor__participante__persona__id', 'tutor__participante__persona__apellido1', 'tutor__participante__persona__apellido2', 'tutor__participante__persona__nombres', 'carrera__id', 'carrera__nombre', 'carrera__coordinacion__id', 'carrera__coordinacion__nombre').filter(carrera__coordinacion__id=codifacu, periodo_id=idper, status=True, tutor__status=True).distinct().order_by('carrera__nombre', 'tutor__participante__persona__apellido1')
                    if codicarr > 0:
                        tematicas = ComplexivoTematica.objects.values_list('tutor__participante__persona__id', 'tutor__participante__persona__apellido1', 'tutor__participante__persona__apellido2', 'tutor__participante__persona__nombres', 'carrera__id', 'carrera__nombre', 'carrera__coordinacion__id', 'carrera__coordinacion__nombre').filter(carrera__id=codicarr, periodo_id=idper, status=True, tutor__status=True).distinct().order_by('carrera__nombre', 'tutor__participante__persona__apellido1')

                    totalhorasgrupo = 0
                    nombrecriterio = ""
                    row_num = 1

                    for tematica in tematicas:
                        horasactividad = 0
                        iddistri = 0
                        idactivi = 0
                        detallehoras = 0

                        if DetalleDistributivo.objects.filter(distributivo__profesor__persona__id=tematica[0], distributivo__periodo=periodo, criteriodocenciaperiodo__criterio__id__in=[29, 99], status=True):
                            distri = DetalleDistributivo.objects.get(distributivo__profesor__persona__id=tematica[0], distributivo__periodo=periodo, criteriodocenciaperiodo__criterio__id__in=[29, 99], status=True)
                            horasactividad = distri.horas
                            nombrecriterio = distri.criteriodocenciaperiodo.criterio.nombre
                            iddistri = distri.id
                            if ActividadDetalleDistributivo.objects.filter(criterio=iddistri, status=True).values_list('id', flat=True):
                                activi = ActividadDetalleDistributivo.objects.get(criterio=iddistri, status=True)
                                idactivi = activi.id
                                if ActividadDetalleDistributivoCarrera.objects.filter(actividaddetalle=idactivi, carrera=tematica[4], status=True):
                                    detalledistri = ActividadDetalleDistributivoCarrera.objects.get(actividaddetalle=idactivi, carrera=tematica[4], status=True)
                                    detallehoras = detalledistri.horas

                        totalhoras = ComplexivoGrupoTematica.objects.filter(status=True, tematica__carrera_id=tematica[4], tematica__periodo_id=idper, activo=True, tematica__tutor__participante__persona_id=tematica[0])
                        totalhorasgrupo = totalhorasgrupo + totalhoras.count()
                        if horasactividad > 0:
                            totalh = int(totalhoras.count())
                            listamatriz.append([tematica[0], tematica[1], tematica[2], tematica[3], tematica[4], tematica[5], totalh, tematica[6], tematica[7], int(horasactividad), int(detallehoras), row_num])
                            row_num += 1
                    data['listacoordinaciones'] = Coordinacion.objects.filter(pk__in=tematicas.values_list('carrera__coordinacion__id', flat=True)).distinct()
                    listacarreras = None

                    if codicarr > 0:
                        listacarreras = Carrera.objects.filter(pk__in=tematicas.values_list('carrera__id', flat=True)).distinct()

                    data['nombrecriterio'] = nombrecriterio
                    data['listacarreras'] = listacarreras
                    data['tematicas'] = listamatriz
                    data['sede'] = Sede.objects.get(pk=1)
                    data['periodo'] = periodo
                    data['totalhorasgrupo'] = totalhorasgrupo
                    return conviert_html_to_pdf(
                        'adm_complexivotematica/reportetutoresgrupo_pdf_antiguos.html',
                        {
                            'pagesize': 'A4',
                            'data': data,
                        }
                    )
                except Exception as ex:
                    pass

            elif action == 'addfirma':
                try:
                    f = FirmasPeriodoGrupoTitulacionForm(request.POST)
                    if f.is_valid():
                        periodogrupo = PeriodoGrupoTitulacion.objects.get(pk=int(request.POST['id']))
                        firmasgrupoperiodo = FirmaPeriodoGrupoTitulacion(periodogrupo=periodogrupo,
                                                                         persona_id=f.cleaned_data['persona'],
                                                                         tipofirma=f.cleaned_data['tipofirma'])
                        firmasgrupoperiodo.save(request)
                        return JsonResponse({"result": "ok"})
                    else:
                        raise NameError('Error')
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

            elif action == 'deletefirma':
                try:
                    firmaperiodotitulacion = FirmaPeriodoGrupoTitulacion.objects.get(pk=int(request.POST['codfirma']))
                    firmaperiodotitulacion.delete()
                    return JsonResponse({"result": "ok"})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": u"Error al vincular rubro."})

            elif action == 'deletetematica':
                try:
                    tematica = ComplexivoTematica.objects.get(pk=int(request.POST['codtematica']))
                    tematica.delete()
                    return JsonResponse({"result": "ok"})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": u"Error al vincular rubro."})

            elif action == 'editartutor':
                try:
                    f = ParticipanteTematicaForm(request.POST)
                    if f.is_valid():
                        grupo = ComplexivoGrupoTematica.objects.get(id=request.POST['id'])
                        grupo.activo = False
                        grupo.save(request)

                        participante = ParticipanteGrupoInvestigacion.objects.get(id=f.cleaned_data['participante'].id)

                        participante_tematica = ParticipanteTematica.objects.get(id=grupo.tematica.tutor_id)

                        participante_tematica.participante = participante
                        participante_tematica.pk = None
                        participante_tematica.save(request)

                        complexivo_tematica = ComplexivoTematica.objects.get(id=grupo.tematica_id)
                        complexivo_tematica.pk = None
                        complexivo_tematica.tutor = participante_tematica
                        complexivo_tematica.save(request)

                        grupo.tematica = complexivo_tematica
                        grupo.pk = None
                        grupo.activo = True
                        grupo.duplicado = True
                        grupo.save(request)

                        for dg in ComplexivoDetalleGrupo.objects.filter(status=True, grupo_id=request.POST['id']):
                            dg.grupo = grupo
                            dg.pk = None
                            dg.save(request)

                        for p in ComplexivoPropuestaPractica.objects.filter(status=True, grupo_id=request.POST['id']):
                            propuestaantigua = ComplexivoPropuestaPractica.objects.get(status=True, grupo_id=request.POST['id'], id=p.id)
                            p.grupo = grupo
                            p.pk = None
                            p.save(request)
                            for arch in ComplexivoPropuestaPracticaArchivo.objects.filter(status=True, propuesta=propuestaantigua):
                                arch.propuesta = p
                                arch.pk = None
                                arch.save()

                        #
                        # for a in ComplexivoAcompanamiento.objects.filter(status=True,grupo_id=request.POST['id']):
                        #     a.grupo=grupo
                        #     a.pk=None
                        #     a.save(request)

                        log(u"Modificar tutor : %s" % grupo, request, "edit")
                        return JsonResponse({"result": "ok"})
                    else:
                        raise NameError('Error')
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

            elif action == 'listadofechatribunal':
                try:
                    fecha = convertir_fecha(request.POST['fecha'])
                    presidente = request.POST['presidente']
                    secretario = request.POST['secretario']
                    delegado = request.POST['delegado']
                    moderador = request.POST['moderador']
                    idhora = request.POST['idhora']
                    resultado = periodo.listadofechatribunal(fecha, presidente, secretario, delegado, moderador)
                    grupo = ComplexivoGrupoTematica.objects.get(pk=int(request.POST['idgrupo']))
                    data['lista'] = resultado
                    listabloqueo = []
                    data['turnotitulacion'] = turnotitulacion = TurnoTitulacion.objects.filter(status=True).order_by('id')
                    fecha = '2022-05-05'
                    for resul in resultado:
                        if resul[2]:
                            resultadopresidente = resul[2].split(',')
                            hora1 = resultadopresidente[0]
                            horacomienza = fecha + ' ' + hora1
                            date_time_objcomienza = datetime.strptime(horacomienza, '%Y-%m-%d %H:%M:%S')
                            hcomienza = date_time_objcomienza.time()
                            hora2 = resultadopresidente[1]
                            horatermina = fecha + ' ' + hora2
                            date_time_objtermina = datetime.strptime(horatermina, '%Y-%m-%d %H:%M:%S')
                            htermina = date_time_objtermina.time()
                            for lturno in turnotitulacion:
                                if lturno.comienza == resul[5]:
                                    listabloqueo.append(lturno.id)
                                if lturno.comienza >= hcomienza and lturno.comienza <= htermina:
                                    listabloqueo.append(lturno.id)
                                if lturno.termina >= hcomienza and lturno.termina <= htermina:
                                    listabloqueo.append(lturno.id)
                        if resul[3]:
                            resultadopresidente = resul[3].split(',')
                            hora1 = resultadopresidente[0]
                            horacomienza = fecha + ' ' + hora1
                            date_time_objcomienza = datetime.strptime(horacomienza, '%Y-%m-%d %H:%M:%S')
                            hcomienza = date_time_objcomienza.time()
                            hora2 = resultadopresidente[1]
                            horatermina = fecha + ' ' + hora2
                            date_time_objtermina = datetime.strptime(horatermina, '%Y-%m-%d %H:%M:%S')
                            htermina = date_time_objtermina.time()
                            for lturno in turnotitulacion:
                                if lturno.comienza >= hcomienza and lturno.comienza <= htermina:
                                    listabloqueo.append(lturno.id)
                                if lturno.termina >= hcomienza and lturno.termina <= htermina:
                                    listabloqueo.append(lturno.id)
                        if resul[4]:
                            resultadopresidente = resul[4].split(',')
                            hora1 = resultadopresidente[0]
                            horacomienza = fecha + ' ' + hora1
                            date_time_objcomienza = datetime.strptime(horacomienza, '%Y-%m-%d %H:%M:%S')
                            hcomienza = date_time_objcomienza.time()
                            hora2 = resultadopresidente[1]
                            horatermina = fecha + ' ' + hora2
                            date_time_objtermina = datetime.strptime(horatermina, '%Y-%m-%d %H:%M:%S')
                            htermina = date_time_objtermina.time()
                            for lturno in turnotitulacion:
                                if lturno.comienza >= hcomienza and lturno.comienza <= htermina:
                                    listabloqueo.append(lturno.id)
                                if lturno.termina >= hcomienza and lturno.termina <= htermina:
                                    listabloqueo.append(lturno.id)

                        for lturno in turnotitulacion:
                            if resul[5]:
                                # if resul[5] != grupo.horadefensa:
                                if grupo.id != resul[9]:
                                    if lturno.comienza == resul[5]:
                                        listabloqueo.append(lturno.id)
                            if resul[6]:
                                # if resul[6] != grupo.horadefensa:
                                if grupo.id != resul[10]:
                                    if lturno.comienza == resul[6]:
                                        listabloqueo.append(lturno.id)
                            if resul[7]:
                                if grupo.id != resul[11]:
                                    if lturno.comienza == resul[7]:
                                        listabloqueo.append(lturno.id)

                    data['listabloqueo'] = listabloqueo
                    data['idhora'] = idhora
                    template = get_template("adm_complexivotematica/listadofechatribunal.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", "listado": json_content})
                except Exception as ex:
                    pass

            elif action == 'addtiempo':
                try:
                    with transaction.atomic():

                        form = ConfiguracionComplexivoHabilitaPropuestaForm(request.POST)
                        if form.is_valid():
                            instance = ConfiguracionComplexivoHabilitaPropuesta(
                                diasalumno=form.cleaned_data['diasalumno'],
                                diasdocente=form.cleaned_data['diasdocente'],
                                activo=form.cleaned_data['activo'])
                            instance.save(request)
                            log(u'Adicionó tiempo : %s' % instance, request, "add")
                            return JsonResponse({"result": False}, safe=False)
                        else:
                            return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()],
                                                 "message": "Error en el formulario"})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

            elif action == 'edittiempo':
                try:
                    with transaction.atomic():
                        instance = ConfiguracionComplexivoHabilitaPropuesta.objects.get(pk=request.POST['id'])
                        form = ConfiguracionComplexivoHabilitaPropuestaForm(request.POST)
                        if form.is_valid():
                            instance.diasalumno = form.cleaned_data['diasalumno']
                            instance.diasdocente = form.cleaned_data['diasdocente']
                            instance.activo = form.cleaned_data['activo']
                            instance.save(request)
                            log(u'Editó tiempo : %s' % instance, request, "edit")
                            return JsonResponse({"result": False}, safe=False)
                        else:
                            return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()],
                                                 "message": "Error en el formulario"})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

            elif action == 'estadoactivorequisitomateria':
                try:
                    with transaction.atomic():
                        requisitomateria = RequisitoMateriaUnidadIntegracionCurricular.objects.get(pk=int(request.POST['id']))
                        if requisitomateria.activo:
                            requisitomateria.activo = False
                        else:
                            requisitomateria.activo = True
                        requisitomateria.save(request)
                        log(u'cambio estado activo materia titulacion: %s' % requisitomateria, request, "estadoactivar")
                        res_json = {"error": False}
                except Exception as ex:
                    res_json = {'error': True, "message": "Error: {}".format(ex)}
                return JsonResponse(res_json, safe=False)

            elif action == 'estadoinscripcionrequisitomateria':
                try:
                    with transaction.atomic():
                        requisitomateria = RequisitoMateriaUnidadIntegracionCurricular.objects.get(pk=int(request.POST['id']))
                        if requisitomateria.inscripcion:
                            requisitomateria.inscripcion = False
                        else:
                            requisitomateria.inscripcion = True
                        requisitomateria.save(request)
                        log(u'cambio estado inscripcion materia titulacion: %s' % requisitomateria, request, "estadoactivar")
                        res_json = {"error": False}
                except Exception as ex:
                    res_json = {'error': True, "message": "Error: {}".format(ex)}
                return JsonResponse(res_json, safe=False)

            elif action == 'estadotitulacionrequisitomateria':
                try:
                    with transaction.atomic():
                        requisitomateria = RequisitoMateriaUnidadIntegracionCurricular.objects.get(pk=int(request.POST['id']))
                        if requisitomateria.titulacion:
                            requisitomateria.titulacion = False
                        else:
                            requisitomateria.titulacion = True
                        requisitomateria.save(request)
                        log(u'cambio estado titulacion materia titulacion: %s' % requisitomateria, request, "estadoactivar")
                        res_json = {"error": False}
                except Exception as ex:
                    res_json = {'error': True, "message": "Error: {}".format(ex)}
                return JsonResponse(res_json, safe=False)

            elif action == 'listadorequisitostitulacion':
                try:
                    lista = []
                    idmate = int(request.POST['idmate'])
                    idmalla = int(request.POST['idmalla'])
                    listadorequisitostitulacion = RequisitoTitulacionMalla.objects.filter(malla_id=idmalla, status=True).exclude(requisito_id__in=RequisitoMateriaUnidadIntegracionCurricular.objects.values_list('requisito_id').filter(materia_id=idmate, status=True)).order_by('id')
                    for lis in listadorequisitostitulacion:
                        lista.append([lis.requisito.id, lis.requisito.nombre])
                    data = {"results": "ok", 'listado': lista}
                    return JsonResponse(data)
                except Exception as ex:
                    pass

            elif action == 'listadomateriastitulacionmalla':
                try:
                    lista = []
                    idcarrera = int(request.POST['idcarrera'])
                    listaprodecesoratitulacion = AsignaturaMalla.objects.filter(malla__carrera_id=idcarrera, status=True)
                    for lis in listaprodecesoratitulacion:
                        lista.append([lis.id, lis.asignatura.nombre])
                    data = {"results": "ok", 'listado': lista}
                    return JsonResponse(data)
                except Exception as ex:
                    pass

            elif action == 'adicionarrequisitomateria':
                try:
                    idmateria = request.POST['idmateria']
                    lista = request.POST['lista'].split(',')
                    for elemento in lista:
                        if not RequisitoMateriaUnidadIntegracionCurricular.objects.filter(requisito_id=elemento, materia_id=idmateria, status=True):
                            requisitomateria = RequisitoMateriaUnidadIntegracionCurricular(requisito_id=elemento,
                                                                                           materia_id=idmateria,
                                                                                           titulacion=True)
                            requisitomateria.save(request)
                    return JsonResponse({'result': 'ok'})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

            elif action == 'configuratiporubrica':
                try:
                    grupo = GrupoTitulacionIC.objects.get(pk=request.POST['codgrupo'])
                    grupo.tiporubrica = request.POST['opc']
                    grupo.save(request)
                    return JsonResponse({'result': 'ok'})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

            elif action == 'addmateriatitulacionrubrica':
                try:
                    materia = Materia.objects.get(pk=request.POST['idmateria'])
                    lista = request.POST['lista'].split(',')
                    if not materia.grupotitulacionic_set.filter(status=True):
                        grupo = GrupoTitulacionIC(materia=materia)
                        grupo.save(request)
                    else:
                        grupo = materia.grupotitulacionic_set.filter(status=True)[0]
                    numorden = 0
                    for elemento in lista:
                        numorden += 1
                        if not MateriaGrupoTitulacion.objects.filter(grupo=grupo, asignaturamalla_id=elemento, status=True):
                            materiagrupo = MateriaGrupoTitulacion(grupo=grupo,
                                                                  asignaturamalla_id=elemento,
                                                                  nombre='',
                                                                  puntaje=0,
                                                                  orden=numorden)
                            materiagrupo.save(request)
                    return JsonResponse({'result': 'ok'})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

            if action == 'deleterubricatitulacion':
                try:
                    rubricamateria = MateriaGrupoTitulacion.objects.get(pk=request.POST['id'])
                    if MateriaGrupoTitulacion.objects.filter(grupo_id=request.POST['idcodigrupo']).count()==1:
                        grupo = GrupoTitulacionIC.objects.get(pk=request.POST['idcodigrupo'])
                        grupo.delete()
                    else:
                        rubricamateria.delete()
                    return JsonResponse({"result": "ok"})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

            elif action == 'generaractamasivo':
                try:
                    id_mate = request.POST['id_mate']
                    lista = request.POST['lista'].split(',')
                    grupotit = GrupoTitulacionIC.objects.get(materia_id=id_mate, status=True)
                    if not grupotit.grupofirma_set.values('id').filter(status=True, activo=True).exists():
                        return JsonResponse({"result": "bad", "mensaje": u"No tiene grupo activo de firma."})
                    grupofirmaactiva = grupotit.grupofirma_set.filter(status=True, activo=True)[0]

                    if not grupofirmaactiva.firmagrupotitulacion_set.values('id').filter(status=True).exists():
                        return JsonResponse({"result": "bad", "mensaje": u"Favor configurar docentes a firmar acta de titulación."})
                    for elemento in lista:
                        asignado = MateriaTitulacion.objects.get(pk=elemento)
                        asignado.numeroacta = MateriaTitulacion.objects.filter(status=True).order_by('-numeroacta')[0].numeroacta + 1
                        asignado.actacerrada = True
                        asignado.notafinal = asignado.materiaasignada.notafinal
                        asignado.fechaacta = datetime.now().date()
                        asignado.numeromemo = request.POST['id_memo']
                        asignado.grupofirma = grupofirmaactiva
                        asignado.save(request)
                    return JsonResponse({'result': 'ok'})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

            elif action == 'generaractamasivorubrica':
                try:
                    id_mate = request.POST['id_mate']
                    lista = request.POST['lista'].split(',')
                    grupotit = GrupoTitulacionIC.objects.get(materia_id=id_mate, status=True)
                    if not grupotit.grupofirma_set.values('id').filter(status=True, activo=True).exists():
                        return JsonResponse({"result": "bad", "mensaje": u"No tiene grupo activo de firma."})
                    grupofirmaactiva = grupotit.grupofirma_set.filter(status=True, activo=True)[0]

                    if not grupofirmaactiva.firmagrupotitulacion_set.values('id').filter(status=True).exists():
                        return JsonResponse({"result": "bad", "mensaje": u"Favor configurar docentes a firmar acta de titulación."})

                    for elemento in lista:
                        asignado = MateriaTitulacion.objects.get(pk=elemento)
                        asignado.numeroacta = MateriaTitulacion.objects.filter(status=True).order_by('-numeroacta')[0].numeroacta + 1
                        asignado.actacerrada = True
                        asignado.fechaacta = datetime.now().date()
                        asignado.numeromemo = request.POST['id_memo']
                        asignado.grupofirma = grupofirmaactiva
                        asignado.save(request)

                        sumatoria = 0
                        listadomateriagrupo = MateriaGrupoTitulacion.objects.filter(grupo__materia_id=asignado.materiaasignada.materia.id).order_by('orden')
                        for materiagrupo in listadomateriagrupo:
                            listaprodecesoratitulacion = MateriaAsignada.objects.filter(matricula__inscripcion_id=asignado.materiaasignada.matricula.inscripcion.id, materia__asignaturamalla_id=materiagrupo.asignaturamalla.id, status=True).order_by('-id')
                            if not listaprodecesoratitulacion and asignado.materiaasignada.materia.asignaturamalla.malla.modalidad.id == 3:
                                listaprodecesoratitulacion = MateriaAsignada.objects.filter(
                                    matricula__inscripcion_id=asignado.materiaasignada.matricula.inscripcion.id,
                                    materia__asignaturamalla__asignatura_id=materiagrupo.asignaturamalla.asignatura.id,
                                    status=True).order_by('-id')

                                if MateriaAsignada.objects.values("id").filter(
                                        matricula__inscripcion_id=asignado.materiaasignada.matricula.inscripcion.id,
                                        materia__asignaturamalla__asignatura_id=materiagrupo.asignaturamalla.asignatura.id,
                                        status=True):
                                    calculanota = round((listaprodecesoratitulacion[0].notafinal * materiagrupo.puntaje) / 100, 0)
                                    sumatoria = sumatoria + calculanota
                                    if MateriaTitulacionNota.objects.values("id").filter(materiatitulacion=asignado, materiagrupo=materiagrupo).exists():
                                        notarubricaalumno = MateriaTitulacionNota.objects.filter(materiatitulacion=asignado, materiagrupo=materiagrupo)[0]
                                        notarubricaalumno.notafinal = calculanota
                                        notarubricaalumno.save(request)
                                    else:
                                        notarubricaalumno = MateriaTitulacionNota(materiatitulacion=asignado,
                                                                                  materiagrupo=materiagrupo,
                                                                                  notafinal=calculanota)
                                        notarubricaalumno.save(request)

                            else:
                                if MateriaAsignada.objects.values("id").filter(matricula__inscripcion_id=asignado.materiaasignada.matricula.inscripcion.id, materia__asignaturamalla_id=materiagrupo.asignaturamalla.id, status=True):
                                    calculanota = round((listaprodecesoratitulacion[0].notafinal * materiagrupo.puntaje) / 100, 0)
                                    sumatoria = sumatoria + calculanota
                                    if MateriaTitulacionNota.objects.values("id").filter(materiatitulacion=asignado, materiagrupo=materiagrupo).exists():
                                        notarubricaalumno = MateriaTitulacionNota.objects.filter(materiatitulacion=asignado, materiagrupo=materiagrupo)[0]
                                        notarubricaalumno.notafinal = calculanota
                                        notarubricaalumno.save(request)
                                    else:
                                        notarubricaalumno = MateriaTitulacionNota(materiatitulacion=asignado,
                                                                                  materiagrupo=materiagrupo,
                                                                                  notafinal=calculanota)
                                        notarubricaalumno.save(request)
                        asignado.notafinal = sumatoria
                        asignado.save(request)

                    return JsonResponse({'result': 'ok'})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

            elif action == 'verdetallerequisitos':
                try:
                    asignada = MateriaAsignada.objects.get(pk=request.POST['idasignada'])
                    htmlrequisitos = listadovalidarequisitos(asignada.matricula.inscripcion, asignada.materia, True)
                    return htmlrequisitos
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

            elif action == 'actualizanombrerubrica':
                try:
                    modelorubrica = MateriaGrupoTitulacion.objects.get(pk=request.POST['iddetalle'])
                    opc = int(request.POST['opc'])
                    if opc == 1:
                        valortexto = request.POST['valortexto']
                        modelorubrica.nombre = valortexto
                    if opc == 2:
                        valortexto = request.POST['valortexto']
                        modelorubrica.puntaje = valortexto
                    if opc == 3:
                        valortexto = request.POST['valortexto']
                        modelorubrica.orden = valortexto
                    modelorubrica.save()
                    return JsonResponse({"result": "ok"})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": u"Error al actualizar."})

            elif action == 'actualizacampofirmadocente':
                try:
                    firma = FirmaGrupoTitulacion.objects.get(pk=request.POST['iddetalle'])
                    opc = int(request.POST['opc'])
                    if opc == 1:
                        valortexto = request.POST['valortexto']
                        firma.orden = valortexto
                    firma.save()
                    return JsonResponse({"result": "ok"})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": u"Error al actualizar."})

            elif action == 'adddocentefirma':
                try:
                    f = FirmaGrupoTitulacionForm(request.POST)
                    if f.is_valid():
                        grupo = FirmaGrupoTitulacion.objects.filter(grupofirma_id=int(request.POST['id']), status=True).order_by('-orden')
                        if grupo.count() == 0:
                            orden = 1
                        else:
                            orden = grupo[0].orden + 1
                        firmagrupo = FirmaGrupoTitulacion(grupofirma_id=int(request.POST['id']),
                                                          profesor_id=f.cleaned_data['profesor'],
                                                          orden=orden)
                        firmagrupo.save(request)
                        return JsonResponse({"result": "ok"})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": "Error al guardar los datos."})

            elif action == 'deletedocentefirma':
                try:
                    docentefirma = FirmaGrupoTitulacion.objects.get(pk=request.POST['id'])
                    docentefirma.delete()
                    return JsonResponse({"result": "ok"})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

            elif action == 'generargrupofirma':
                try:
                    listado = GrupoFirma.objects.filter(grupo_id=request.POST['id'])
                    for lgrupo in listado:
                        lgrupo.activo=False
                        lgrupo.save(request)
                    grupo = GrupoFirma(grupo_id=request.POST['id'], activo=True)
                    grupo.save(request)
                    return JsonResponse({"result": "ok"})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": u"Error al generar grupo."})

            elif action == 'deletegrupofirma':
                try:
                    grupof = GrupoFirma.objects.get(pk=request.POST['id'])
                    grupof.delete()
                    return JsonResponse({"result": "ok"})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

            elif action == 'nivelesmalla':
                try:
                    lista = []
                    listarequisitos = []
                    listadoniveles = AsignaturaMalla.objects.values_list('nivelmalla_id', 'nivelmalla__nombre').filter(malla_id=int(encrypt(request.POST['id_malla']))).distinct()
                    requisitosmalla = RequisitoTitulacionMalla.objects.filter(malla_id=int(encrypt(request.POST['id_malla'])), status=True).exclude(requisito_id__in=[8,12,13]).order_by('requisito__nombre')
                    for nivel in listadoniveles:
                        lista.append([nivel[0], nivel[1]])
                    for requi in requisitosmalla:
                        listarequisitos.append([requi.id, requi.requisito.nombre])
                    return JsonResponse({'result': 'ok', 'lista': lista, 'listarequisitos': listarequisitos})
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

            elif action == 'estadoobligatoriorequisitomateria':
                try:
                    with transaction.atomic():
                        requisitomateria = RequisitoMateriaUnidadIntegracionCurricular.objects.get(
                            pk=int(request.POST['id']))
                        if requisitomateria.obligatorio:
                            requisitomateria.obligatorio = False
                        else:
                            requisitomateria.obligatorio = True
                        requisitomateria.save(request)
                        log(u'cambio estado obligatorio materia titulacion: %s' % requisitomateria, request,
                            "estadoactivar")
                        res_json = {"error": False}
                except Exception as ex:
                    res_json = {'error': True, "message": "Error: {}".format(ex)}
                return JsonResponse(res_json, safe=False)

            elif action == 'generarplanificacion':
                try:
                    periodo_id = periodo.id
                    if not 'archivo' in request.FILES:
                        raise NameError('Carge un archivo excel para ejecutar la acción.')
                    archivo = request.FILES['archivo']
                    idmat = request.POST['idmat']
                    materia = Materia.objects.get(pk=int(encrypt(idmat)))
                    listado=MateriaTitulacion.objects.filter(materiaasignada__materia=materia, materiaasignada__status=True,
                                                     status=True, materiaasignada__retiramateria=False).order_by(
                        'materiaasignada__matricula__inscripcion__persona__apellido1',
                        'materiaasignada__matricula__inscripcion__persona__apellido2',
                        'materiaasignada__matricula__inscripcion__persona__nombres')
                    detallemodeloevaluativo_id = DetalleModeloEvaluativo.objects.filter(
                        modelo=materia.modeloevaluativo).filter(
                        Q(alternativa_id=31) | Q(alternativa_id=146)
                    ).values_list('id', flat=True).first()
                    workbook = openpyxl.load_workbook(archivo)
                    sheet = workbook.worksheets[0]

                    rows = sheet.rows
                    linea = 0
                    for row in rows:
                        linea += 1
                        if linea > 1:
                            if row[1].value:
                                if row[12].value.replace(" ", "") == 'VIRTUAL':
                                    sede_id = 11
                                    aula_id = 289
                                else:
                                    sede_id = 1
                                    if isinstance(row[13].value, int):
                                        aula = row[13].value
                                    else:
                                        aula = int(row[13].value)

                                    aula_id = LaboratorioVirtual.objects.filter(pk=aula).values_list('id', flat=True).first()
                                    if not aula_id:
                                        res_json = {"result": "bad", "message": U"Verifique el aula ingresada alumno %s" % row[2].value}
                                        return JsonResponse(res_json, safe=False)
                                matasignada = None
                                id_materia = None
                                for item in listado:
                                    if isinstance(row[1].value, int):
                                        ced = row[1].value
                                    else:
                                        ced = row[1].value.strip().replace(" ", "")

                                    if item.materiaasignada.matricula.inscripcion.persona.cedula == ced:
                                        id_matricula = item.materiaasignada.matricula.id
                                        id_materia = item.materiaasignada.materia.id
                                        matasignada = item.materiaasignada
                                fecha = row[3].value
                                if not id_materia:
                                    continue

                                if not DisertacionMateriaAsignadaPlanificacion.objects.filter(materiaasignada = matasignada).first():
                                    try:
                                        eDisertacionFechaPlanificacion = DisertacionFechaPlanificacion.objects.get(fecha=fecha,
                                                                                                                   sede_id=sede_id,
                                                                                                                   periodo_id=periodo_id)
                                    except ObjectDoesNotExist:
                                        eDisertacionFechaPlanificacion = DisertacionFechaPlanificacion(fecha=fecha,
                                                                                                       sede_id=sede_id,
                                                                                                       periodo_id=periodo_id)
                                        eDisertacionFechaPlanificacion.save()
                                    horainicio = row[4].value
                                    horafin = row[5].value

                                    eDisertacionTurnoPlanificacion, created = DisertacionTurnoPlanificacion.objects.get_or_create(
                                        horainicio=horainicio,
                                        horafin=horafin,
                                        fechaplanificacion=eDisertacionFechaPlanificacion)

                                    eDisertacionAulaPlanificacion, created = DisertacionAulaPlanificacion.objects.get_or_create(
                                        turnoplanificacion=eDisertacionTurnoPlanificacion, aula_id=aula_id)

                                    grupo = row[11].value
                                    if row[10].value:
                                        responsable = row[10].value if isinstance(row[10].value, int) else row[10].value.strip().replace(" ", "")
                                    else:
                                        responsable = row[6].value if isinstance(row[6].value, int) else row[6].value.strip().replace(" ", "")
                                    eResponsable = None
                                    try:
                                        eResponsable = Persona.objects.get(Q(cedula=responsable) | Q(pasaporte=responsable))
                                    except ObjectDoesNotExist:
                                        eResponsable = None
                                    try:
                                        eMateria = Materia.objects.get(pk=id_materia)
                                    except ObjectDoesNotExist:
                                        eMateria = None
                                    if grupo and eResponsable and eMateria:
                                        try:
                                            eDisertacionGrupoPlanificacion = DisertacionGrupoPlanificacion.objects.get(
                                                aulaplanificacion=eDisertacionAulaPlanificacion, grupo=grupo, materia=eMateria,
                                                detallemodeloevaluativo_id=detallemodeloevaluativo_id)
                                        except ObjectDoesNotExist:
                                            eDisertacionGrupoPlanificacion = DisertacionGrupoPlanificacion(
                                                aulaplanificacion=eDisertacionAulaPlanificacion,
                                                materia=eMateria,
                                                detallemodeloevaluativo_id=detallemodeloevaluativo_id,
                                                grupo=grupo)
                                        eDisertacionGrupoPlanificacion.responsable = eResponsable
                                        eDisertacionGrupoPlanificacion.save()
                                        if not (row[6].value == "" or row[6].value == None or row[6].value == ''):
                                            profesor_1 = row[6].value.strip().replace(" ", "") if row[6].value else ""
                                            if eProfesor1 := Persona.objects.filter(Q(cedula=profesor_1) | Q(pasaporte=profesor_1)).first():
                                                eDisertacionTribunalPlanificacion1, created = DisertacionTribunalPlanificacion.objects.get_or_create(
                                                    grupoplanificacion=eDisertacionGrupoPlanificacion, responsable=eProfesor1)

                                        if not (row[7].value == "" or row[7].value == None or row[7].value == ''):
                                            profesor_2 = row[7].value.strip().replace(" ", "") if row[7].value else ""
                                            if eProfesor2 := Persona.objects.filter(Q(cedula=profesor_2) | Q(pasaporte=profesor_2)).first():
                                                eDisertacionTribunalPlanificacion2, created = DisertacionTribunalPlanificacion.objects.get_or_create(
                                                    grupoplanificacion=eDisertacionGrupoPlanificacion,
                                                    responsable=eProfesor2)

                                        if not (row[8].value == "" or row[8].value == None or row[8].value == ''):
                                            profesor_3 = row[8].value.strip().replace(" ", "") if row[8].value else ""
                                            if eProfesor3 := Persona.objects.filter(Q(cedula=profesor_3) | Q(pasaporte=profesor_3)).first():
                                                eDisertacionTribunalPlanificacion3, created = DisertacionTribunalPlanificacion.objects.get_or_create(
                                                    grupoplanificacion=eDisertacionGrupoPlanificacion,
                                                    responsable=eProfesor3)

                                        if not (row[9].value == "" or row[9].value == None or row[9].value == ''):
                                            profesor_4 = row[9].value.strip().replace(" ", "") if row[9].value else ""
                                            if eProfesor4 := Persona.objects.filter(Q(cedula=profesor_4) | Q(pasaporte=profesor_4)).first():
                                                eDisertacionTribunalPlanificacion3, created = DisertacionTribunalPlanificacion.objects.get_or_create(
                                                    grupoplanificacion=eDisertacionGrupoPlanificacion,
                                                    responsable=eProfesor4)

                                        try:
                                            eMatricula = Matricula.objects.get(pk=id_matricula)
                                        except ObjectDoesNotExist:
                                            eMatricula = None
                                        if eMatricula:
                                            try:
                                                eMateriaAsignada = MateriaAsignada.objects.get(matricula=eMatricula,
                                                                                               materia=eMateria)
                                            except ObjectDoesNotExist:
                                                eMateriaAsignada = None
                                            if eMateriaAsignada:
                                                try:
                                                    eDisertacionMateriaAsignadaPlanificacion = DisertacionMateriaAsignadaPlanificacion.objects.get(
                                                        grupoplanificacion=eDisertacionGrupoPlanificacion,
                                                        materiaasignada=eMateriaAsignada)
                                                except ObjectDoesNotExist:
                                                    eDisertacionMateriaAsignadaPlanificacion = DisertacionMateriaAsignadaPlanificacion(
                                                        grupoplanificacion=eDisertacionGrupoPlanificacion,
                                                        materiaasignada=eMateriaAsignada)
                                                    eDisertacionMateriaAsignadaPlanificacion.save()
                    log(u'Cargo Archivo disertacion de examenes: %s. Este archivo fue cargado: %s' % (persona,request.FILES['archivo'].name), request, "add")
                    res_json = {"result": "ok", "message": "Planificación generada."}
                    return JsonResponse(res_json, safe=False)
                except Exception as ex:
                    res_json = {"result": "bad", "message": U"Error al obtener los datos. %s" % ex }
                return JsonResponse(res_json,safe=False)

            elif action == 'adddisertacionalumno':
                try:
                    periodo_id = periodo.id
                    idmat = request.POST['id']
                    cdla = request.POST['cdla']
                    materia = Materia.objects.get(pk=int(encrypt(idmat)))
                    listado=MateriaTitulacion.objects.filter(materiaasignada__materia=materia, materiaasignada__status=True,
                                                     status=True, materiaasignada__retiramateria=False).order_by(
                        'materiaasignada__matricula__inscripcion__persona__apellido1',
                        'materiaasignada__matricula__inscripcion__persona__apellido2',
                        'materiaasignada__matricula__inscripcion__persona__nombres')

                    detallemodeloevaluativo_id = DetalleModeloEvaluativo.objects.filter(modelo=materia.modeloevaluativo).filter(
                        Q(alternativa_id=31) | Q(alternativa_id=146)).values_list('id', flat=True).first()

                    if request.POST['sede'] == '0':
                        res_json = {"result": "bad", "message": U"Seleccione una sede."}
                        return JsonResponse(res_json, safe=False)

                    if request.POST['sede'] == '11':
                        sede_id = 11
                        aula_id = 289
                    else:
                        sede_id = 1
                        aula_id = LaboratorioVirtual.objects.filter(pk=request.POST['aula']).values_list('id', flat=True).first()

                    for item in listado:
                        if item.materiaasignada.matricula.inscripcion_id == int(encrypt(cdla)):
                            id_matricula = item.materiaasignada.matricula.id
                            id_materia = item.materiaasignada.materia.id
                    fecha_str = request.POST['fecha']
                    fecha_obj = datetime.strptime(fecha_str, '%d-%m-%Y')
                    fecha = fecha_obj.strftime('%Y-%m-%d')
                    try:
                        eDisertacionFechaPlanificacion = DisertacionFechaPlanificacion.objects.get(fecha=fecha,
                                                                                                   sede_id=sede_id,
                                                                                                   periodo_id=periodo_id)
                    except ObjectDoesNotExist:
                        eDisertacionFechaPlanificacion = DisertacionFechaPlanificacion(fecha=fecha,
                                                                                       sede_id=sede_id,
                                                                                       periodo_id=periodo_id)
                        eDisertacionFechaPlanificacion.save()
                    horainicio = request.POST['horainicio']
                    horafin = request.POST['horafin']
                    try:
                        eDisertacionTurnoPlanificacion = DisertacionTurnoPlanificacion.objects.get(
                            fechaplanificacion=eDisertacionFechaPlanificacion, horainicio=horainicio,
                            horafin=horafin)
                    except ObjectDoesNotExist:
                        eDisertacionTurnoPlanificacion = DisertacionTurnoPlanificacion(
                            fechaplanificacion=eDisertacionFechaPlanificacion,
                            horainicio=horainicio,
                            horafin=horafin)
                        eDisertacionTurnoPlanificacion.save()
                    try:
                        eDisertacionAulaPlanificacion = DisertacionAulaPlanificacion.objects.get(
                            turnoplanificacion=eDisertacionTurnoPlanificacion, aula_id=aula_id)
                    except ObjectDoesNotExist:
                        eDisertacionAulaPlanificacion = DisertacionAulaPlanificacion(
                            turnoplanificacion=eDisertacionTurnoPlanificacion,
                            aula_id=aula_id)
                        eDisertacionAulaPlanificacion.save()

                    grupo = request.POST['grupo']
                    responsable = request.POST['profesor1']
                    eResponsable = None
                    try:
                        eResponsable = Persona.objects.get(pk=responsable)
                    except ObjectDoesNotExist:
                        eResponsable = None
                    try:
                        eMateria = Materia.objects.get(pk=id_materia)
                    except ObjectDoesNotExist:
                        eMateria = None
                    if grupo and eResponsable and eMateria:
                        try:
                            eDisertacionGrupoPlanificacion = DisertacionGrupoPlanificacion.objects.get(
                                aulaplanificacion=eDisertacionAulaPlanificacion, grupo=grupo, materia=eMateria,
                                detallemodeloevaluativo_id=detallemodeloevaluativo_id)
                        except ObjectDoesNotExist:
                            eDisertacionGrupoPlanificacion = DisertacionGrupoPlanificacion(
                                aulaplanificacion=eDisertacionAulaPlanificacion,
                                materia=eMateria,
                                detallemodeloevaluativo_id=detallemodeloevaluativo_id,
                                grupo=grupo)
                        eDisertacionGrupoPlanificacion.responsable = eResponsable
                        eDisertacionGrupoPlanificacion.save()
                        profesor_1 =  request.POST.get('profesor1', None)
                        if profesor_1:
                            try:
                                eProfesor1 = Persona.objects.get(pk=profesor_1)
                            except ObjectDoesNotExist:
                                eProfesor1 = None
                            if eProfesor1:
                                try:
                                    eDisertacionTribunalPlanificacion1 = DisertacionTribunalPlanificacion.objects.get(
                                        grupoplanificacion=eDisertacionGrupoPlanificacion,
                                        responsable=eProfesor1)
                                except ObjectDoesNotExist:
                                    eDisertacionTribunalPlanificacion1 = DisertacionTribunalPlanificacion(
                                        grupoplanificacion=eDisertacionGrupoPlanificacion,
                                        responsable=eProfesor1)
                                    eDisertacionTribunalPlanificacion1.save()
                        profesor_2 =  request.POST.get('profesor2', None)
                        if profesor_2:
                            try:
                                eProfesor2 = Persona.objects.get(pk=profesor_2)
                            except ObjectDoesNotExist:
                                eProfesor2 = None
                            if eProfesor2:
                                try:
                                    eDisertacionTribunalPlanificacion2 = DisertacionTribunalPlanificacion.objects.get(
                                        grupoplanificacion=eDisertacionGrupoPlanificacion,
                                        responsable=eProfesor2)
                                except ObjectDoesNotExist:
                                    eDisertacionTribunalPlanificacion2 = DisertacionTribunalPlanificacion(
                                        grupoplanificacion=eDisertacionGrupoPlanificacion,
                                        responsable=eProfesor2)
                                    eDisertacionTribunalPlanificacion2.save()
                        profesor_3 = request.POST.get('profesor3', None)
                        if profesor_3:
                            try:
                                eProfesor3 = Persona.objects.get(pk=profesor_3)
                            except ObjectDoesNotExist:
                                eProfesor3 = None
                            if eProfesor3:
                                try:
                                    eDisertacionTribunalPlanificacion3 = DisertacionTribunalPlanificacion.objects.get(
                                        grupoplanificacion=eDisertacionGrupoPlanificacion,
                                        responsable=eProfesor3)
                                except ObjectDoesNotExist:
                                    eDisertacionTribunalPlanificacion3 = DisertacionTribunalPlanificacion(
                                        grupoplanificacion=eDisertacionGrupoPlanificacion,
                                        responsable=eProfesor3)
                                    eDisertacionTribunalPlanificacion3.save()
                        profesor_4 =  request.POST.get('profesor4', None)
                        if profesor_4:
                            try:
                                eProfesor4 = Persona.objects.get(pk=profesor_4)
                            except ObjectDoesNotExist:
                                eProfesor4 = None
                            if eProfesor4:
                                try:
                                    eDisertacionTribunalPlanificacion4 = DisertacionTribunalPlanificacion.objects.get(
                                        grupoplanificacion=eDisertacionGrupoPlanificacion,
                                        responsable=eProfesor4)
                                except ObjectDoesNotExist:
                                    eDisertacionTribunalPlanificacion4 = DisertacionTribunalPlanificacion(
                                        grupoplanificacion=eDisertacionGrupoPlanificacion,
                                        responsable=eProfesor4)
                                    eDisertacionTribunalPlanificacion4.save()

                        try:
                            eMatricula = Matricula.objects.get(pk=id_matricula)
                        except ObjectDoesNotExist:
                            eMatricula = None
                        if eMatricula:
                            try:
                                eMateriaAsignada = MateriaAsignada.objects.get(matricula=eMatricula,
                                                                               materia=eMateria)
                            except ObjectDoesNotExist:
                                eMateriaAsignada = None
                            if eMateriaAsignada:
                                try:
                                    eDisertacionMateriaAsignadaPlanificacion = DisertacionMateriaAsignadaPlanificacion.objects.get(
                                        grupoplanificacion=eDisertacionGrupoPlanificacion,
                                        materiaasignada=eMateriaAsignada)
                                except ObjectDoesNotExist:
                                    eDisertacionMateriaAsignadaPlanificacion = DisertacionMateriaAsignadaPlanificacion(
                                        grupoplanificacion=eDisertacionGrupoPlanificacion,
                                        materiaasignada=eMateriaAsignada)
                                    eDisertacionMateriaAsignadaPlanificacion.save()

                    log(u'Creo disertacion de examenes: %s. del alumno: %s.' % (persona,eMateriaAsignada.matricula.inscripcion.persona), request, "add")
                    res_json = {"result": "ok", "message": "Planificación generada."}
                    return JsonResponse(res_json, safe=False)
                except Exception as ex:
                    res_json = {"result": "bad", "message": U"Error al obtener los datos"}
                return JsonResponse(res_json,safe=False)

            elif action == 'verdetalledisertacion':
                try:
                    data['asignada']= asignada = MateriaAsignada.objects.filter(pk=request.POST['idasignada']).first()
                    detallemodeloevaluativo_id = DetalleModeloEvaluativo.objects.filter(
                        modelo=asignada.materia.modeloevaluativo).filter(Q(alternativa_id=31) | Q(alternativa_id=146)
                    ).values_list('id',flat=True).first()

                    disermatasignada = DisertacionMateriaAsignadaPlanificacion.objects.filter(materiaasignada = asignada).first()
                    if not disermatasignada:
                        transaction.set_rollback(True)
                        return JsonResponse({"result": "bad", "mensaje": u"No tiene asignada disertacion"})

                    if asignada.matricula.inscripcion.mi_malla():
                        mimalla = asignada.matricula.inscripcion.mi_malla()
                        ultimonivelmalla = asignada.matricula.inscripcion.mi_malla().ultimo_nivel_malla()
                        penultimonivel = asignada.matricula.inscripcion.mi_malla().ultimo_nivel_malla().orden - 1
                    data['ultimonivelmalla'] = ultimonivelmalla
                    data['penultimonivel'] = penultimonivel
                    data['grupoplanificacion'] = grupoplanificacion = DisertacionGrupoPlanificacion.objects.filter(pk=disermatasignada.grupoplanificacion.id).first()
                    data['tibunal'] = tribunal = DisertacionTribunalPlanificacion.objects.filter(grupoplanificacion=grupoplanificacion).order_by('fecha_creacion')
                    template = get_template("adm_complexivotematica/modal/detalledisertacion.html")
                    return JsonResponse({'result': "ok", 'data': template.render(data)})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

            elif action == 'aditdisertacionalumno':
                try:
                    if request.POST['sede'] == '0':
                        res_json = {"result": "bad", "message": U"Seleccione una sede."}
                        return JsonResponse(res_json, safe=False)

                    fecha_str = request.POST['fecha']
                    fecha_obj = datetime.strptime(fecha_str, '%d-%m-%Y')
                    fecha = fecha_obj.strftime('%Y-%m-%d')

                    if request.POST['sede'] == '11':
                        sede_id = 11
                        aula_id = 289
                    else:
                        sede_id = 1
                        aula_id = LaboratorioVirtual.objects.filter(pk=request.POST['aula']).values_list('id',flat=True).first()

                    disermatasignada = DisertacionMateriaAsignadaPlanificacion.objects.get( materiaasignada_id=int(encrypt(request.POST['id'])))

                    detallemodeloevaluativo_id = DetalleModeloEvaluativo.objects.filter(modelo=disermatasignada.materiaasignada.materia.modeloevaluativo).filter(Q(alternativa_id=31) | Q(alternativa_id=146)).values_list('id', flat=True).first()
                    fechaplanificacion, created = DisertacionFechaPlanificacion.objects.get_or_create(
                        fecha=fecha,
                        sede_id=sede_id,
                        periodo_id=periodo.id)

                    turnoplanificacion, created = DisertacionTurnoPlanificacion.objects.get_or_create(
                        horainicio=request.POST['horainicio'],
                        horafin=request.POST['horafin'],
                        fechaplanificacion=fechaplanificacion)

                    eDisertacionAulaPlanificacion, created = DisertacionAulaPlanificacion.objects.get_or_create(
                        turnoplanificacion=turnoplanificacion,aula_id=aula_id)

                    eResponsable = Persona.objects.get(pk=request.POST['profesor1'])
                    if grupoplanificacion := DisertacionGrupoPlanificacion.objects.filter(pk=disermatasignada.grupoplanificacion_id, grupo=request.POST['grupo']).first():
                        grupoplanificacion.responsable = eResponsable
                        grupoplanificacion.aulaplanificacion = eDisertacionAulaPlanificacion
                        grupoplanificacion.save()
                        tribunal = DisertacionTribunalPlanificacion.objects.filter(
                            grupoplanificacion=grupoplanificacion).delete()
                    else:
                        grupoplanificacion, created = DisertacionGrupoPlanificacion.objects.get_or_create(aulaplanificacion=eDisertacionAulaPlanificacion,
                            materia=disermatasignada.materiaasignada.materia,
                            detallemodeloevaluativo_id=detallemodeloevaluativo_id,
                            grupo=request.POST['grupo'],
                            responsable=eResponsable
                        )
                        disermatasignada.grupoplanificacion = grupoplanificacion
                        disermatasignada.save()

                    if profesor_1 := request.POST.get('profesor1', None):
                        if eProfesor1 := Persona.objects.get(pk=profesor_1):
                            eDisertacionTribunalPlanificacion1, created = DisertacionTribunalPlanificacion.objects.get_or_create(
                                grupoplanificacion=grupoplanificacion,responsable=eProfesor1)

                    if profesor_2 := request.POST.get('profesor2', None):
                        if eProfesor2 := Persona.objects.get(pk=profesor_2):
                            eDisertacionTribunalPlanificacion2, created = DisertacionTribunalPlanificacion.objects.get_or_create(
                                grupoplanificacion=grupoplanificacion,responsable=eProfesor2)

                    if profesor_3 := request.POST.get('profesor3', None):
                        if eProfesor3 := Persona.objects.get(pk=profesor_3):
                            eDisertacionTribunalPlanificacion3 = DisertacionTribunalPlanificacion.objects.get_or_create(
                                grupoplanificacion=grupoplanificacion,responsable=eProfesor3)

                    if profesor_4 := request.POST.get('profesor4', None):
                        if eProfesor4 := Persona.objects.get(pk=profesor_4):
                            eDisertacionTribunalPlanificacion4 = DisertacionTribunalPlanificacion.objects.get_or_create(
                                grupoplanificacion=grupoplanificacion,responsable=eProfesor4)

                    log(u'Edito disertacion de examenes: %s. del alumno: %s.' % (
                    persona, disermatasignada.materiaasignada.matricula.inscripcion.persona), request, "edit")
                    return JsonResponse({"result": "ok", "message": u"Registros actualizados"})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

            elif action == 'deletedisertacion':
                try:
                    asignada = MateriaAsignada.objects.filter(pk=request.POST['id']).first()
                    disermatasignada = DisertacionMateriaAsignadaPlanificacion.objects.filter(
                        materiaasignada=asignada).first()
                    grupoplanificacion = DisertacionGrupoPlanificacion.objects.filter(
                        pk=disermatasignada.grupoplanificacion.id).first()
                    tribunal = DisertacionTribunalPlanificacion.objects.filter(
                        grupoplanificacion=grupoplanificacion).order_by('fecha_creacion')
                    aulaplanificacion = DisertacionAulaPlanificacion.objects.get(pk=grupoplanificacion.aulaplanificacion.id)
                    turnoplanificacion = DisertacionTurnoPlanificacion.objects.get(pk=aulaplanificacion.turnoplanificacion.id)
                    disermatasignada.delete()
                    if not DisertacionMateriaAsignadaPlanificacion.objects.filter(grupoplanificacion=grupoplanificacion).count() > 1:
                        grupoplanificacion.delete()
                    #tribunal.delete()
                    #aulaplanificacion.delete()
                    #turnoplanificacion.delete()
                    log(u'Elimino disertacion de examenes: %s. del alumno: %s.' % (persona,asignada.matricula.inscripcion.persona), request, "del")
                    return JsonResponse({"result": "ok"})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})



        return JsonResponse({"result": "bad", "mensaje": u"Solicitud Incorrecta."})
    else:
        if 'action' in request.GET:
            data['action'] = action = request.GET['action']

            if action == 'tiempopropuesta':
                try:
                    data['title'] = 'Tiempos para habilitar propuesta'
                    data['listado'] = listado = ConfiguracionComplexivoHabilitaPropuesta.objects.filter(status=True)
                    return render(request, 'adm_complexivotematica/viewtiempop.html', data)
                except Exception as ex:
                    pass

            if action == 'asignaturastitulacion':
                try:
                    data['title'] = 'Asignaturas de titulación'
                    search, filtro, url_vars = request.GET.get('s', ''), Q(status=True), ''
                    idc = None
                    listado1 = Materia.objects.filter(asignaturamalla__validarequisitograduacion=True, nivel__periodo=periodo, asignaturamalla__status=True, status=True)
                    listado2 = Materia.objects.filter(asignaturamalla__malla_id=383, nivel__periodo=periodo, asignaturamalla__status=True, status=True)
                    listado = listado1 | listado2
                    data['listadocarreras'] = Carrera.objects.filter(id__in=listado.values_list('asignaturamalla__malla__carrera_id', flat=True)).distinct()
                    if GrupoTitulacionIC.objects.filter(materia__nivel__periodo=periodo, status=True).count() < listado.count():
                        for lmate in listado:
                            adicionoagrupofirma(request,lmate)

                    if 's' in request.GET:
                        search = request.GET['s'].strip()
                        grupo = GrupoTitulacionIC.objects.filter(Q(materia__asignaturamalla__asignatura__nombre__icontains=search), materia__nivel__periodo=periodo, status=True).annotate(totalrubricas = Count('materiagrupotitulacion__id')).order_by('materia__asignaturamalla__malla__carrera__nombre', 'materia__paralelo')
                    else:
                        if 'idc' in request.GET:
                            idc = int(request.GET['idc'])
                            if idc > 0:
                                grupo = GrupoTitulacionIC.objects.filter(materia__nivel__periodo=periodo, materia__asignaturamalla__malla__carrera_id=idc, status=True).annotate(totalrubricas=Count('materiagrupotitulacion__id')).order_by('materia__asignaturamalla__malla__carrera__nombre', 'materia__paralelo')
                            else:
                                grupo = GrupoTitulacionIC.objects.filter(materia__nivel__periodo=periodo, status=True).annotate(totalrubricas = Count('materiagrupotitulacion__id')).order_by('materia__asignaturamalla__malla__carrera__nombre', 'materia__paralelo')
                        else:
                            grupo = GrupoTitulacionIC.objects.filter(materia__nivel__periodo=periodo, status=True).annotate(totalrubricas = Count('materiagrupotitulacion__id')).order_by('materia__asignaturamalla__malla__carrera__nombre', 'materia__paralelo')

                    paging = MiPaginador(grupo, 25)
                    p = 1
                    url_vars += "&action=asignaturastitulacion"
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['idc'] = idc if idc else ""
                    data['search'] = search if search else ""
                    data["url_vars"] = url_vars
                    data['listado'] = page.object_list
                    return render(request, 'adm_complexivotematica/asignaturastitulacion.html', data)
                except Exception as ex:
                    pass

            if action == 'listadoalumnos':
                try:
                    data['title'] = 'Asignaturas de titulación'
                    search, filtro, url_vars = request.GET.get('s', ''), Q(status=True), ''
                    data['materia'] = materia = Materia.objects.get(pk=int(encrypt(request.GET['idm'])))
                    adicionoagrupofirma(request, materia)
                    data['sinrequisitos'] =  sinrequisitos = not materia.requisitomateriaunidadintegracioncurricular_set.values("id").filter(status=True).exists()
                    # if not materia.requisitomateriaunidadintegracioncurricular_set.values("id").filter(status=True).exists():
                    #     sinrequisitos = True
                    # data['sinrequisitos'] = sinrequisitos
                    hoy = datetime.now().date()
                    rezagados = materia.asignaturamalla.malla.id == 383
                    # if materia.asignaturamalla.malla.id == 383:
                    #     rezagados = True
                    graduados = MateriaTitulacion.objects.values("id").filter(materiaasignada__materia=materia, estadograduado=True).exists()
                    if not graduados and not sinrequisitos:
                        response = llenar_requisitostitulacion(materia, request)
                        if response['resp'] == 'error':
                            return HttpResponseRedirect("/adm_complexivotematica?info={}".format(response['msg']))
                        if materia.materiaasignada_set.filter(status=True).count() > MateriaTitulacion.objects.values("id").filter(materiaasignada__materia=materia, materiaasignada__status=True).count():
                            lmateriaasignada = materia.materiaasignada_set.filter(status=True)
                            for lasignado in lmateriaasignada:
                                if not MateriaTitulacion.objects.values("id").filter(materiaasignada=lasignado).exists():
                                    mtitulacion = MateriaTitulacion(materiaasignada=lasignado,
                                                                    rezagados=rezagados)
                                    mtitulacion.save()
                    if 's' in request.GET:
                        search = request.GET['s'].strip()
                        ss = search.split(' ')
                        if len(ss) == 1:
                            listado = MateriaTitulacion.objects.filter(Q(materiaasignada__matricula__inscripcion__persona__nombres__icontains=search) |
                                                                       Q(materiaasignada__matricula__inscripcion__persona__apellido1__icontains=search) |
                                                                       Q(materiaasignada__matricula__inscripcion__persona__apellido2__icontains=search) |
                                                                       Q(materiaasignada__matricula__inscripcion__persona__cedula__icontains=search) |
                                                                       Q(materiaasignada__matricula__inscripcion__persona__pasaporte__icontains=search) |
                                                                       Q(materiaasignada__matricula__inscripcion__persona__usuario__username__icontains=search),
                                                                       materiaasignada__materia=materia, materiaasignada__status=True, status=True,materiaasignada__retiramateria=False).order_by('materiaasignada__matricula__inscripcion__persona__apellido1', 'materiaasignada__matricula__inscripcion__persona__apellido2', 'materiaasignada__matricula__inscripcion__persona__nombres')

                        else:
                            listado = MateriaTitulacion.objects.filter(Q(materiaasignada__matricula__inscripcion__persona__apellido1__icontains=ss[0]) &
                                                                       Q(materiaasignada__matricula__inscripcion__persona__apellido2__icontains=ss[1]), materiaasignada__materia=materia, materiaasignada__status=True, status=True, materiaasignada__retiramateria=False).order_by('materiaasignada__matricula__inscripcion__persona__apellido1', 'materiaasignada__matricula__inscripcion__persona__apellido2', 'materiaasignada__matricula__inscripcion__persona__nombres')
                    else:
                        listado = MateriaTitulacion.objects.filter(materiaasignada__materia=materia, materiaasignada__status=True, status=True, materiaasignada__retiramateria=False).order_by('materiaasignada__matricula__inscripcion__persona__apellido1', 'materiaasignada__matricula__inscripcion__persona__apellido2', 'materiaasignada__matricula__inscripcion__persona__nombres')
                    data['cronograma'] = materia.cronogramacalificaciones()
                    numerofilas = 25
                    paging = MiPaginador(listado, numerofilas)
                    p = 1
                    url_vars += "&action=listadoalumnos&idm="+request.GET['idm']
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                            if p == 1:
                                numerofilasguiente = numerofilas
                            else:
                                numerofilasguiente = numerofilas * (p - 1)
                        else:
                            p = paginasesion
                            if p == 1:
                                numerofilasguiente = numerofilas
                            else:
                                numerofilasguiente = numerofilas * (p - 1)
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['numerofilasguiente'] = numerofilasguiente
                    data['numeropagina'] = p
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data["url_vars"] = url_vars
                    data['listado'] = page.object_list
                    return render(request, 'adm_complexivotematica/listadoalumnos.html', data)
                except Exception as ex:
                    pass

            if action == 'listadoalumnosrubrica':
                try:
                    data['title'] = 'Asignaturas de titulación'
                    search, filtro, url_vars = request.GET.get('s', ''), Q(status=True), ''
                    data['materia'] = materia = Materia.objects.get(pk=int(encrypt(request.GET['idm'])))
                    data['tienerubricas'] = tienerubricas = MateriaGrupoTitulacion.objects.values('id').filter(grupo__materia_id=materia.id, status=True).exists()
                    # if not MateriaGrupoTitulacion.objects.values('id').filter(grupo__materia_id=materia.id, status=True).exists():
                    #     tienerubricas = False
                    # data['tienerubricas'] = tienerubricas
                    data['sinrequisitos'] = sinrequisitos = not materia.requisitomateriaunidadintegracioncurricular_set.values("id").filter(status=True).exists()
                    # if not materia.requisitomateriaunidadintegracioncurricular_set.values("id").filter(status=True).exists():
                    #     sinrequisitos = True
                    # data['sinrequisitos'] = sinrequisitos
                    hoy = datetime.now().date()
                    rezagados = materia.asignaturamalla.malla.id == 383
                    # if materia.asignaturamalla.malla.id == 383:
                    #     rezagados = True
                    graduados = MateriaTitulacion.objects.values("id").filter(materiaasignada__materia=materia, estadograduado=True).exists()
                    # if not graduados and sinrequisitos:
                    if not graduados:
                        response = llenar_requisitostitulacion(materia, request)
                        if response['resp'] == 'error':
                            return HttpResponseRedirect("/adm_complexivotematica?info={}".format(response['msg']))
                        if materia.materiaasignada_set.filter(status=True).count() > MateriaTitulacion.objects.values("id").filter(materiaasignada__materia=materia, materiaasignada__status=True).count():
                            lmateriaasignada = materia.materiaasignada_set.filter(status=True)
                            for lasignado in lmateriaasignada:
                                if not MateriaTitulacion.objects.values("id").filter(materiaasignada=lasignado).exists():
                                    mtitulacion = MateriaTitulacion(materiaasignada=lasignado,
                                                                    rezagados=rezagados)
                                    mtitulacion.save()
                    if 's' in request.GET:
                        search = request.GET['s'].strip()
                        ss = search.split(' ')
                        if len(ss) == 1:
                            listado = MateriaTitulacion.objects.filter(Q(materiaasignada__matricula__inscripcion__persona__nombres__icontains=search) |
                                                                       Q(materiaasignada__matricula__inscripcion__persona__apellido1__icontains=search) |
                                                                       Q(materiaasignada__matricula__inscripcion__persona__apellido2__icontains=search) |
                                                                       Q(materiaasignada__matricula__inscripcion__persona__cedula__icontains=search) |
                                                                       Q(materiaasignada__matricula__inscripcion__persona__pasaporte__icontains=search) |
                                                                       Q(materiaasignada__matricula__inscripcion__persona__usuario__username__icontains=search),
                                                                       materiaasignada__materia=materia, materiaasignada__status=True, status=True,materiaasignada__retiramateria=False).order_by('materiaasignada__matricula__inscripcion__persona__apellido1', 'materiaasignada__matricula__inscripcion__persona__apellido2', 'materiaasignada__matricula__inscripcion__persona__nombres')

                        else:
                            listado = MateriaTitulacion.objects.filter(Q(materiaasignada__matricula__inscripcion__persona__apellido1__icontains=ss[0]) &
                                                                       Q(materiaasignada__matricula__inscripcion__persona__apellido2__icontains=ss[1]), materiaasignada__materia=materia, materiaasignada__status=True, status=True, materiaasignada__retiramateria=False).order_by('materiaasignada__matricula__inscripcion__persona__apellido1', 'materiaasignada__matricula__inscripcion__persona__apellido2', 'materiaasignada__matricula__inscripcion__persona__nombres')
                    else:
                        listado = MateriaTitulacion.objects.filter(materiaasignada__materia=materia, materiaasignada__status=True, status=True, materiaasignada__retiramateria=False).order_by('materiaasignada__matricula__inscripcion__persona__apellido1', 'materiaasignada__matricula__inscripcion__persona__apellido2', 'materiaasignada__matricula__inscripcion__persona__nombres')
                    data['cronograma'] = materia.cronogramacalificaciones()
                    numerofilas = 25
                    paging = MiPaginador(listado, numerofilas)
                    p = 1
                    url_vars += "&action=listadoalumnosrubrica&idm="+request.GET['idm']
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                            if p == 1:
                                numerofilasguiente = numerofilas
                            else:
                                numerofilasguiente = numerofilas * (p - 1)
                        else:
                            p = paginasesion
                            if p == 1:
                                numerofilasguiente = numerofilas
                            else:
                                numerofilasguiente = numerofilas * (p - 1)
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['numerofilasguiente'] = numerofilasguiente
                    data['numeropagina'] = p
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data["url_vars"] = url_vars
                    data['listado'] = page.object_list
                    data['listadorubricas'] = MateriaGrupoTitulacion.objects.filter(grupo__materia=materia, status=True).order_by('orden')
                    return render(request, 'adm_complexivotematica/listadoalumnosrubrica.html', data)
                except Exception as ex:
                    pass

            if action == 'actacalificaciones':
                try:
                    data['title'] = 'Acta de calificación'
                    search, filtro, url_vars = request.GET.get('s', ''), Q(status=True), ''
                    data['materia'] = materia = Materia.objects.get(pk=int(encrypt(request.GET['idm'])))
                    sinrequisitos = False
                    if not materia.requisitomateriaunidadintegracioncurricular_set.values("id").filter(status=True).exists():
                        sinrequisitos = True
                    data['sinrequisitos'] = sinrequisitos
                    hoy = datetime.now().date()
                    rezagados = False
                    if materia.asignaturamalla.malla.id == 383:
                        rezagados = True
                    if 's' in request.GET:
                        search = request.GET['s'].strip()
                        ss = search.split(' ')
                        if len(ss) == 1:
                            listado = MateriaAsignada.objects.filter(Q(matricula__inscripcion__persona__nombres__icontains=search) |
                                                                     Q(matricula__inscripcion__persona__apellido1__icontains=search) |
                                                                     Q(matricula__inscripcion__persona__apellido2__icontains=search) |
                                                                     Q(matricula__inscripcion__persona__cedula__icontains=search) |
                                                                     Q(matricula__inscripcion__persona__pasaporte__icontains=search) |
                                                                     Q(matricula__inscripcion__persona__usuario__username__icontains=search),
                                                                     materia=materia, status=True, retiramateria=False).order_by('matricula__inscripcion__persona__apellido1', 'matricula__inscripcion__persona__apellido2', 'matricula__inscripcion__persona__nombres')

                        else:
                            listado = MateriaAsignada.objects.filter(Q(matricula__inscripcion__persona__apellido1__icontains=ss[0]) &
                                                                     Q(matricula__inscripcion__persona__apellido2__icontains=ss[1]), materia=materia, status=True, retiramateria=False).order_by('matricula__inscripcion__persona__apellido1', 'matricula__inscripcion__persona__apellido2', 'matricula__inscripcion__persona__nombres')
                    else:
                        listado = MateriaAsignada.objects.filter(materia=materia, status=True, retiramateria=False).order_by('matricula__inscripcion__persona__apellido1', 'matricula__inscripcion__persona__apellido2', 'matricula__inscripcion__persona__nombres')
                    data['cronograma'] = materia.cronogramacalificaciones()
                    numerofilas = 25
                    paging = MiPaginador(listado, numerofilas)
                    p = 1
                    url_vars += "&action=actacalificaciones&idm="+request.GET['idm']
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                            if p == 1:
                                numerofilasguiente = numerofilas
                            else:
                                numerofilasguiente = numerofilas * (p - 1)
                        else:
                            p = paginasesion
                            if p == 1:
                                numerofilasguiente = numerofilas
                            else:
                                numerofilasguiente = numerofilas * (p - 1)
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['numerofilasguiente'] = numerofilasguiente
                    data['numeropagina'] = p
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data["url_vars"] = url_vars
                    data['listado'] = page.object_list
                    return render(request, 'adm_complexivotematica/actacalificaciones.html', data)
                except Exception as ex:
                    pass

            if action == 'detallerequisitos':
                try:
                    data['inscripcion'] = materiaasignado = MateriaAsignada.objects.get(pk=int(encrypt(request.GET['idasignado'])))
                    data['listarequisitos'] = materiaasignado.materia.requisitomateriaunidadintegracioncurricular_set.filter(activo=True, titulacion=True, status=True)
                    return render(request, 'adm_complexivotematica/detallerequisitos.html', data)
                except Exception as ex:
                    pass

            elif action == 'asignacupogrupo':
                try:
                    data['title'] = u'Asignar Cupos a Grupos'
                    data['idt'] = int(request.GET['idt'])

                    tematica = ComplexivoTematica.objects.get(pk=int(request.GET['idt']))

                    data['permite_modificar'] = True
                    data['tematica'] = tematica

                    grupos = tematica.complexivotematicagrupocupo_set.filter(status=True).order_by('numerogrupo')
                    lista_grupos = [[g.id, g.numerogrupo, g.cupoasignado, 'S' if g.enuso else 'N'] for g in grupos]
                    data['grupos'] = lista_grupos
                    template = get_template("adm_complexivotematica/asignacupogrupotematica.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content, 'title': data['title']})
                except Exception as ex:
                    pass

            elif action == 'novisibletematica':
                try:
                    tematica = Tematica.objects.get(pk=request.GET['id'])

                    if ComplexivoTematica.objects.filter(tematica=tematica, carrera=miscarreras[0]).exists():
                        comtematica = ComplexivoTematica.objects.filter(tematica=tematica, carrera=miscarreras[0])
                        for com in comtematica:
                            if com.tiene_grupos():
                                return
                        ComplexivoTematica.objects.filter(tematica=tematica, carrera=miscarreras[0]).delete()
                    return HttpResponseRedirect('/adm_complexivotematica?action=listtematicas')
                except Exception as ex:
                    pass

            elif action == 'visibletematica':
                try:
                    tematica = Tematica.objects.get(pk=request.GET['id'])
                    participantes = tematica.participantetematica_set.filter(status=True, estutor=True)
                    for participante in participantes:
                        complexivotematica = ComplexivoTematica(tematica=tematica, tutor=participante, director=profesor, carrera=miscarreras[0])
                        complexivotematica.save(request)
                    return HttpResponseRedirect('/adm_complexivotematica?action=listtematicas')
                except Exception as ex:
                    transaction.set_rollback(True)
                    pass

            elif action == 'addtematicacupo':
                try:
                    data['title'] = u"Añadir Participante de Investigación"
                    data['tematica'] = tematica = ComplexivoTematica.objects.get(pk=request.GET['id'])
                    form = ComplexivoTematicaForm(initial={
                        'cupo': tematica.cupo,
                        'tutor': tematica.tutor,
                        'maxintegrantes': tematica.maxintegrantes
                    })
                    form.agregar(tematica.tematica)
                    data['form'] = form
                    return render(request, "adm_complexivotematica/addtematicacupo.html", data)
                except Exception as ex:
                    pass

            elif action == 'vergrupos':
                try:
                    data['title'] = u"Grupos Asignados"
                    if 'c' in request.GET:
                        data['c'] = request.GET['c']
                    if 's' in request.GET:
                        data['s'] = request.GET['s']
                    if 'per' in request.GET:
                        data['per'] = request.GET['per']
                    data['tematica'] = tematica = ComplexivoTematica.objects.get(pk=request.GET['id'])
                    data['grupos'] = ComplexivoGrupoTematica.objects.filter(status=True, tematica=tematica, activo=True).order_by('id')
                    data['form2'] = ComplexivoTematicaObservacionForm()
                    return render(request, "adm_complexivotematica/lstgrupo.html", data)
                except Exception as ex:
                    pass

            elif action == 'detallematricula':
                try:
                    data['title'] = u'Incripciones al Proceso de Titulación'
                    data['tematicaid'] = request.GET['idtematica']
                    data['c'] = request.GET['c']
                    matricula = MatriculaTitulacion.objects.get(pk=int(request.GET['idm']))
                    alter = AlternativaTitulacion.objects.get(pk=matricula.alternativa_id)
                    inscripcion = Inscripcion.objects.get(pk=matricula.inscripcion_id)
                    data = valida_matricular_estudiante(data, alter, inscripcion)
                    return render(request, "adm_complexivotematica/detallematriculatitulacion.html", data)
                except Exception as ex:
                    pass

            elif action == 'detailgrupo':
                try:
                    data['grupo'] = ComplexivoGrupoTematica.objects.get(pk=request.GET['id'])
                    template = get_template("adm_complexivotematica/detallegrupo.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content})
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

            elif action == 'detallecalificaciontribunal':
                try:
                    data['participante'] = participante = ComplexivoDetalleGrupo.objects.get(pk=request.GET['id'])
                    data['detallecalificacion'] = detallecalificacion = participante.calificacionrubricatitulacion_set.filter(status=True).order_by('tipojuradocalificador')
                    data['promediopuntajetrabajointegral'] = detallecalificacion.values_list('puntajetrabajointegral').aggregate(promedio=Avg('puntajetrabajointegral'))['promedio']
                    data['promediodefensaoral'] = detallecalificacion.values_list('puntajedefensaoral').aggregate(promedio=Avg('puntajedefensaoral'))['promedio']
                    data['promediofinal'] = detallecalificacion.values_list('puntajerubricas').aggregate(promedio=Avg('puntajerubricas'))['promedio']
                    # if participante.matricula.alternativa.tipotitulacion.rubrica:
                    if participante.rubrica:
                        modelorubricatitulacion = ModeloRubricaTitulacion.objects.filter(rubrica=participante.rubrica, status=True).order_by('id')
                        for detjurado in detallecalificacion:
                            if not detjurado.calificaciondetallemodelorubricatitulacion_set.filter(status=True):
                                for rubmodelo in modelorubricatitulacion:
                                    calificacionmodelorubrica = CalificacionDetalleModeloRubricaTitulacion(calificacionrubrica=detjurado,
                                                                                                           modelorubrica=rubmodelo,
                                                                                                           puntaje=0)
                                    calificacionmodelorubrica.save()
                                    detcalificacion = CalificacionDetalleRubricaTitulacion.objects.filter(calificacionrubrica=detjurado, rubricatitulacion__modelorubrica=calificacionmodelorubrica.modelorubrica, status=True).aggregate(valor=Sum('puntaje'))['valor']
                                    calificacionmodelorubrica.puntaje = detcalificacion
                                    calificacionmodelorubrica.save()
                        data['listadomodelorubrica'] = participante.rubrica.modelorubricatitulacion_set.filter(status=True).order_by('orden')
                        template = get_template("adm_complexivotematica/detallecalificaciontribunalmodrubrica.html")
                    # else:
                    #     template = get_template("adm_complexivotematica/detallecalificaciontribunal.html")

                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content})
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

            elif action == 'detalleproceso':
                try:
                    data['participante'] = participante = ComplexivoDetalleGrupo.objects.get(pk=request.GET['id'])
                    data['propuestas'] = participante.grupo.complexivopropuestapractica_set.filter(status=True).order_by('id')
                    template = get_template("adm_complexivotematica/detalleproceso.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content})
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

            elif action == 'asignartribunal':
                try:
                    # puede_realizar_accion(request, 'sga.puede_editar_cupotematica')
                    data['title'] = u"Asignar Tribunal"
                    if 'c' in request.GET:
                        data['c'] = request.GET['c']
                    data['roles'] = CARGOS_JURADO_SUSTENTACION
                    data['grupo'] = grupo = ComplexivoGrupoTematica.objects.get(pk=int(request.GET['id']))
                    form = ComplexivoTribunalCalificador(initial={'presidente': grupo.presidentepropuesta.id if grupo.presidentepropuesta else 0,
                                                                  'secretario': grupo.secretariopropuesta.id if grupo.secretariopropuesta else 0,
                                                                  'delegado': grupo.delegadopropuesta.id if grupo.delegadopropuesta else 0,
                                                                  'moderador': grupo.moderador.id if grupo.moderador else 0,
                                                                  'fecha': grupo.fechadefensa if grupo.fechadefensa else datetime.now().date(),
                                                                  # 'hora': str(grupo.horadefensa) if grupo.horadefensa else datetime.now().time().strftime('%H:%M'),
                                                                  'hora': str(grupo.horadefensa)[:5] if grupo.horadefensa else '',
                                                                  'lugar': grupo.lugardefensa if grupo.lugardefensa else ''})
                    if grupo.presidentepropuesta:
                        form.fields['presidente'].widget.attrs['descripcion'] = grupo.presidentepropuesta
                        form.fields['presidente'].widget.attrs['value'] = grupo.presidentepropuesta.id
                    else:
                        form.fields['presidente'].widget.attrs['descripcion'] = '----------------'
                        form.fields['presidente'].widget.attrs['value'] = 0
                    if grupo.secretariopropuesta:
                        form.fields['secretario'].widget.attrs['descripcion'] = grupo.secretariopropuesta
                        form.fields['secretario'].widget.attrs['value'] = grupo.secretariopropuesta.id
                    else:
                        form.fields['secretario'].widget.attrs['descripcion'] = '----------------'
                        form.fields['secretario'].widget.attrs['value'] = 0
                    if grupo.delegadopropuesta:
                        form.fields['delegado'].widget.attrs['descripcion'] = grupo.delegadopropuesta
                        form.fields['delegado'].widget.attrs['value'] = grupo.delegadopropuesta.id
                    else:
                        form.fields['delegado'].widget.attrs['descripcion'] = '----------------'
                        form.fields['delegado'].widget.attrs['value'] = 0
                    if grupo.moderador:
                        form.fields['moderador'].widget.attrs['descripcion'] = grupo.moderador
                        form.fields['moderador'].widget.attrs['value'] = grupo.moderador.id
                    else:
                        form.fields['moderador'].widget.attrs['descripcion'] = '----------------'
                        form.fields['moderador'].widget.attrs['value'] = 0
                    data['form'] = form
                    return render(request, "adm_complexivotematica/asignartribunal.html", data)
                except Exception as ex:
                    pass

            elif action == 'deletegrupo':
                try:
                    # puede_realizar_accion(request, 'sga.puede_editar_cupotematica')
                    data['title'] = u'Eliminar Grupo'
                    data['grupo'] = grupo = ComplexivoGrupoTematica.objects.get(pk=request.GET['id'])
                    data['tematica'] = grupo.tematica
                    mensaje = u"Esta seguro(a) de eliminar el grupo conformado por :"
                    x = 0
                    for participante in grupo.participantes():
                        if x == 0:
                            mensaje = mensaje + u" %s" % (participante.matricula)
                        else:
                            mensaje = mensaje + u", %s" % (participante.matricula)
                        x = x + 1
                    data['mensaje'] = mensaje
                    return render(request, "adm_complexivotematica/deletegrupo.html", data)

                except Exception as ex:
                    pass

            elif action == 'deletetribunal':
                try:
                    data['title'] = u'Eliminar Tribunal'
                    data['grupo'] = grupo = ComplexivoGrupoTematica.objects.get(pk=request.GET['id'])
                    data['tematica'] = grupo.tematica
                    mensaje = u"Esta seguro(a) de eliminar el tribunal conformado por :"
                    data['mensaje'] = mensaje
                    return render(request, "adm_complexivotematica/deletetribunal.html", data)

                except Exception as ex:
                    pass

            elif action == 'actualizar_graduado':
                data['title'] = u'Actualizar tribunal en graduados'
                data['grupo'] = ComplexivoGrupoTematica.objects.get(pk=request.GET['id'])
                return render(request, "adm_complexivotematica/actualizardatosgraduados.html", data)

            elif action == 'detalle':
                try:
                    tematica = ComplexivoTematica.objects.get(pk=request.GET['id'])
                    tema = tematica.tematica
                    lista = []
                    for linea in tema.listar_lineas():
                        lista.append(linea.linea.__str__())
                    datos = {
                        'grupo': tema.grupo.__str__(),
                        'director': tema.grupo.director.__str__(),
                        'codirector': tema.grupo.codirector.__str__() if tema.grupo.codirector != None else 'No asignado',
                        'lista': lista
                    }
                    return JsonResponse(datos)
                except Exception as ex:
                    pass

            elif action == 'busqueda':
                try:
                    carrera = Carrera.objects.get(pk=request.GET['c'])
                    lista = ComplexivoDetalleGrupo.objects.values_list('matricula').filter(status=True, matricula__estado__in=[1], matricula__alternativa__grupotitulacion__periodogrupo__abierto=True, grupo__activo=True)
                    q = request.GET['q'].upper().strip()
                    s = q.split(" ")
                    if s.__len__() == 2:
                        # matri = MatriculaTitulacion.objects.filter(alternativa__grupotitulacion__periodogrupo__abierto=True, inscripcion__persona__apellido1__icontains=s[0],inscripcion__persona__apellido2__icontains=s[1],inscripcion__carrera=carrera, estado=1).exclude(pk__in=lista).distinct()[:20]
                        matri = MatriculaTitulacion.objects.filter(alternativa__grupotitulacion__periodogrupo__abierto=True, inscripcion__persona__apellido1__icontains=s[0], inscripcion__persona__apellido2__icontains=s[1], estado=1).exclude(pk__in=lista).distinct()[:20]
                    else:
                        # matri = MatriculaTitulacion.objects.filter(Q(inscripcion__persona__nombres__contains=s[0]) | Q(inscripcion__persona__apellido1__contains=s[0]) | Q(inscripcion__persona__apellido2__contains=s[0]) | Q(inscripcion__persona__cedula__contains=s[0])).filter(alternativa__grupotitulacion__periodogrupo__abierto=True, inscripcion__carrera=carrera, estado=1).exclude(pk__in=lista).distinct()[:20]
                        matri = MatriculaTitulacion.objects.filter(Q(inscripcion__persona__nombres__contains=s[0]) | Q(inscripcion__persona__apellido1__contains=s[0]) | Q(inscripcion__persona__apellido2__contains=s[0]) | Q(inscripcion__persona__cedula__contains=s[0])).filter(alternativa__grupotitulacion__periodogrupo__abierto=True, estado=1).exclude(pk__in=lista).distinct()[:20]
                    data = {"result": "ok", "results": [{"id": x.id, "name": x.flexbox_repr()} for x in matri]}
                    return JsonResponse(data)
                except Exception as ex:
                    pass

            elif action == 'buscarprofesor':
                try:
                    q = request.GET['q'].upper().strip()
                    s = q.split(" ")
                    if s.__len__() == 2:
                        profe = Profesor.objects.filter(persona__apellido1__icontains=s[0], persona__apellido2__icontains=s[1]).distinct()[:20]

                    else:
                        profe = Profesor.objects.filter(Q(persona__nombres__contains=s[0]) | Q(persona__apellido1__contains=s[0]) | Q(persona__apellido2__contains=s[0]) | Q(persona__cedula__contains=s[0])).distinct()[:20]
                    data = {"result": "ok", "results": [{"id": x.id, "name": x.flexbox_repr(), 'nombre': str(x)} for x in profe]}
                    return JsonResponse(data)
                except Exception as ex:
                    pass

            elif action == 'addparticipante':
                try:
                    puede_realizar_accion(request, 'sga.puede_editar_cupotematica')
                    data['grupo'] = grupo = ComplexivoGrupoTematica.objects.get(pk=int(request.GET['id']))
                    data['id'] = grupo.id
                    if grupo.tiene_cupo():
                        data['tematica'] = grupo.tematica
                        data['actione'] = "addparticipante"
                        restantes = grupo.cupo_restante()
                        data['maxintegrante'] = restantes
                        form = InscripcionTitulacionForm()
                        data['idcarrera'] = request.GET['idcarrera']
                        form.fields['participantetitulacion'].queryset = MatriculaTitulacion.objects.none()
                        data['form'] = form
                        template = get_template("adm_complexivotematica/modal/addparticipantegrupo.html")
                        return JsonResponse({"result": True, 'data': template.render(data)})
                    else:
                        return JsonResponse({"result": False, 'message': u"Numero de participantes completos."})
                except Exception as ex:
                    return JsonResponse({"result": False, 'message': u"Error al obtener los datos."})

            elif action == 'delparticipante':
                try:
                    puede_realizar_accion(request, 'sga.puede_editar_cupotematica')
                    data['grupo'] = grupo = ComplexivoGrupoTematica.objects.get(pk=int(request.GET['id']))
                    data['tematica'] = grupo.tematica
                    data['integrantes'] = grupo.complexivodetallegrupo_set.filter(status=True).order_by('matricula__inscripcion__persona__apellido1', 'matricula__inscripcion__persona__apellido2', 'matricula__inscripcion__persona__nombres')
                    template = get_template("adm_complexivotematica/delparticipante.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content})
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

            elif action == 'addgrupo':
                try:
                    tematica = ComplexivoTematica.objects.get(pk=int(request.GET['id']))
                    data['tematica'] = tematica
                    data['id'] = tematica.id
                    data['actione'] = "addgrupo"
                    restantes = 2
                    template = get_template("adm_complexivotematica/addparticipante.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content, 'maxintegrante': restantes})
                    # else:
                    #     return JsonResponse({"result": "bad", "mensaje": u"Numero de participantes completos."})
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

            if action == 'addgrupotematica':
                try:
                    data['title'] = u'Adicionar grupoparticipante'
                    tematica = ComplexivoTematica.objects.get(pk=int(request.GET['id']))
                    data['tematica'] = tematica
                    data['id'] = tematica.id
                    data['actione'] = "addgrupo"
                    form = InscripcionTitulacionForm()
                    data['idcarrera'] = request.GET['idcarrera']
                    form.fields['participantetitulacion'].queryset = MatriculaTitulacion.objects.none()
                    data['form'] = form
                    template = get_template("adm_complexivotematica/modal/addparticipantegrupo.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, 'message': str(ex)})

            if action == 'buscarpersonases':
                try:
                    carrera = Carrera.objects.get(pk=request.GET['c'])
                    lista = ComplexivoDetalleGrupo.objects.values_list('matricula').filter(status=True, matricula__estado__in=[1], matricula__alternativa__grupotitulacion__periodogrupo__abierto=True, grupo__activo=True)
                    q = request.GET['q'].upper().strip()
                    s = q.split(" ")
                    if s.__len__() == 2:
                        # matri = MatriculaTitulacion.objects.filter(alternativa__grupotitulacion__periodogrupo__abierto=True, inscripcion__persona__apellido1__icontains=s[0],inscripcion__persona__apellido2__icontains=s[1],inscripcion__carrera=carrera, estado=1).exclude(pk__in=lista).distinct()[:20]
                        matri = MatriculaTitulacion.objects.filter(alternativa__grupotitulacion__periodogrupo__abierto=True, inscripcion__persona__apellido1__icontains=s[0], inscripcion__persona__apellido2__icontains=s[1], estado=1).exclude(pk__in=lista).distinct()[:20]
                    else:
                        # matri = MatriculaTitulacion.objects.filter(Q(inscripcion__persona__nombres__contains=s[0]) | Q(inscripcion__persona__apellido1__contains=s[0]) | Q(inscripcion__persona__apellido2__contains=s[0]) | Q(inscripcion__persona__cedula__contains=s[0])).filter(alternativa__grupotitulacion__periodogrupo__abierto=True, inscripcion__carrera=carrera, estado=1).exclude(pk__in=lista).distinct()[:20]
                        matri = MatriculaTitulacion.objects.filter(Q(inscripcion__persona__nombres__contains=s[0]) | Q(inscripcion__persona__apellido1__contains=s[0]) | Q(inscripcion__persona__apellido2__contains=s[0]) | Q(inscripcion__persona__cedula__contains=s[0])).filter(alternativa__grupotitulacion__periodogrupo__abierto=True, estado=1).exclude(pk__in=lista).distinct()[:20]
                    data = {"result": "ok", "results": [{"id": x.id, "name": x.flexbox_repr()} for x in matri]}
                    return JsonResponse(data)
                except Exception as ex:
                    pass

            elif action == 'confreporte':
                try:
                    if 'car' in request.GET:
                        carrera = Carrera.objects.get(pk=request.GET['car'])
                    else:
                        carrera = 0
                    periodo = PeriodoGrupoTitulacion.objects.get(pk=request.GET['per'])
                    opcion = request.GET['opcion']
                    respuesta = True

                    if opcion == 'nominatutor':
                        respuesta = ComplexivoDetalleGrupo.objects.filter(status=True, grupo__tematica__carrera=carrera, grupo__tematica__periodo=periodo).exists()

                    if respuesta:
                        return JsonResponse({"result": "ok"})
                    else:
                        return JsonResponse({"result": "bad", "mensaje": "No existen datos para este reporte"})
                except Exception as ex:
                    pass

            elif action == 'nominatematica':
                try:
                    __author__ = 'Unemi'
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('reporte 1')
                    encabezado = 14
                    fil = encabezado + 1
                    col1 = 6
                    tit = 2
                    carrera = 0
                    if 'car' in request.GET:
                        carrera = Carrera.objects.get(pk=request.GET['car'])
                    periodo = PeriodoGrupoTitulacion.objects.get(pk=request.GET['per'])
                    tematicas = ComplexivoTematica.objects.filter(status=True, carrera=carrera, periodo=periodo).distinct('tematica')
                    if not carrera == 0:
                        tit = 1
                        col1 = 5
                        if tematicas.exists():
                            ws.write_merge(7, 7, 0, 1, u'FACULTAD:', stylebnombre)
                            ws.write_merge(8, 8, 0, 1, u'CARRERA:', stylebnombre)
                            ws.write_merge(9, 9, 0, 1, u'DIRECTOR DE CARRERA:', stylebnombre)
                            ws.write(7, 2, u"%s" % carrera.coordinaciones().get().nombre, stylebnombre)
                            ws.merge(7, 7, 2, col1)
                            ws.write(8, 2, str(carrera), stylebnombre)
                            ws.merge(8, 8, 2, col1)
                            ws.write(9, 2, u"%s" % tematicas[0].director, stylebnombre)
                            ws.merge(9, 9, 2, col1)
                    else:
                        tematicas = ComplexivoTematica.objects.filter(status=True, periodo=periodo).distinct('tematica')
                        ws.write(encabezado, 1, 'CARRERA', stylebnotas)
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename =NOMINA DE TEMATICAS ' + str(carrera) + '.xls'

                    if tematicas.exists():
                        ws.col(0).width = 700
                        ws.col(1).width = 6500
                        ws.col(2).width = 6500
                        ws.col(tit).width = 6500
                        ws.col(tit + 1).width = 6500
                        ws.col(tit + 2).width = 5000
                        ws.col(tit + 3).width = 700
                        ws.col(tit + 4).width = 6500
                        ws.write_merge(1, 1, 0, col1, u'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                        ws.write_merge(2, 2, 0, col1, u'VICERRECTORADO ACADÉMICO Y DE INVESTIGACIÓN', title)
                        ws.write_merge(3, 3, 0, col1, u'GESTIÓN TÉCNICA ACADÉMICA', title)
                        ws.write_merge(4, 4, 0, col1, u'PROCESO DE  TITULACIÓN', title)
                        ws.write_merge(5, 5, 0, col1, u'NÓMINA DE LÍNEA DE INVESTIGACIÓN', title)
                        ws.write_merge(10, 11, 0, 1, u'PROCESO DE TITULACIÓN \nGESTIÓN TÉCNICA ACADÉMICA:', stylebnombre)
                        ws.write_merge(10, 11, 2, 3, u'ING. VIVIANA GAIBOR HINOSTROZA, MSC.', stylebnombre)
                        ws.write(6, 0, 'INICIO: ' + str(periodo.fechainicio) + ' FIN: ' + str(periodo.fechafin), subtitle)
                        ws.merge(6, 6, 0, 6)
                        ws.write(encabezado, 0, 'Nº', stylebnotas)
                        ws.write(encabezado, tit, 'LÍNEA DE INVESTIGACIÓN', stylebnotas)
                        ws.write(encabezado, tit + 1, 'GRUPO DE INVESTIGACIÓN', stylebnotas)
                        ws.write(encabezado, tit + 2, 'ÁREA DE INVESTIGACIÓN', stylebnotas)
                        ws.write(encabezado, tit + 3, 'Nº', stylebnotas)
                        ws.write(encabezado, tit + 4, 'ACOMPAÑANTE', stylebnotas)
                        normal.borders = borders
                        i = 1
                        c = 0
                        fil = encabezado + 1
                        for tematica in tematicas:
                            texto = ''
                            acom = ' '
                            filt = fil
                            if tematica.tematica.lineastematica_set.all().exists():
                                for linea in tematica.tematica.lineastematica_set.all():
                                    texto = texto + '\n•' + str(linea.linea)
                            acompanantes = ComplexivoTematica.objects.filter(tematica=tematica.tematica,
                                                                             carrera=tematica.carrera, periodo=periodo,
                                                                             status=True).distinct('tutor')
                            if acompanantes.exists():
                                c = 0
                                for acompanante in acompanantes:
                                    if (acompanante.tematica == tematica.tematica):
                                        c = c + 1
                                        acom = acom + '\n•' + str(acompanante.tutor)

                            ws.write(fil, tit + 2, str(texto), normaliz)
                            ws.write(fil, tit + 4, str(acom), normaliz)
                            ws.write(fil, tit + 3, str(c), normaliz)
                            ws.write(fil, 0, str(i), normal)
                            if carrera == 0:
                                ws.write(fil, 1, u'%s' % tematica.carrera, normal)
                            ws.write(fil, tit, str(tematica.tematica), normal)
                            ws.write(fil, tit + 1, str(tematica.tematica.grupo), normal)
                            fil = filt + 1
                            i = i + 1
                        ws.write(fil + 3, 3, u"_____________________________\nING. VIVIANA GAIBOR,MSC.\nPROCESO DE TITULACIÓN", normalsinborde)
                        ws.merge(fil + 3, fil + 5, 3, 5, normalsinborde)
                        if not carrera == 0:
                            ws.write_merge(fil + 3, fil + 3, 1, 2, u'_____________________________', normalsinborde)
                            ws.write(fil + 4, 1, u"%s" % tematicas[0].director, normalsinborde)
                            ws.merge(fil + 4, fil + 4, 1, 2, normalsinborde)
                            ws.write_merge(fil + 5, fil + 5, 1, 2, u'DIRECTOR(A) DE CARRERA', normalsinborde)
                    wb.save(response)
                    return response

                except Exception as es:
                    pass

            elif action == 'nominaestudiante':
                try:
                    __author__ = 'Unemi'
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('reporte 1')
                    encabezado = 14
                    fil = encabezado + 1
                    col1 = 6
                    tit = 0
                    carrera = 0
                    if 'car' in request.GET:
                        carrera = Carrera.objects.get(pk=request.GET['car'])
                    periodo = PeriodoGrupoTitulacion.objects.get(pk=request.GET['per'])
                    lstid = ComplexivoDetalleGrupo.objects.values_list('id').distinct('grupo').filter(status=True, grupo__tematica__periodo=periodo).exclude(estado=3)
                    grupos = ComplexivoDetalleGrupo.objects.filter(pk__in=lstid).exclude(estado=3)
                    if not carrera == 0:
                        tit = 2
                        col1 = 5
                        grupos = grupos.filter(grupo__tematica__carrera=carrera, grupo__tematica__periodo=periodo).order_by('matricula__inscripcion__persona__apellido1', 'matricula__inscripcion__persona__apellido2')
                        if grupos.exists():
                            ws.write_merge(7, 7, 0, 1, u'FACULTAD:', stylebnombre)
                            ws.write_merge(8, 8, 0, 1, u'CARRERA:', stylebnombre)
                            ws.write_merge(9, 9, 0, 1, u'DIRECTOR DE CARRERA:', stylebnombre)
                            director = u"%s" % (grupos[0].grupo.tematica.director)
                            ws.write(7, 2, u"%s" % carrera.coordinaciones().get().nombre, stylebnombre)
                            ws.merge(7, 7, 2, 6)
                            ws.write(8, 2, str(carrera), stylebnombre)
                            ws.merge(8, 8, 2, 6)
                            ws.write(9, 2, u"%s" % director, stylebnombre)
                            ws.merge(9, 9, 2, 6)
                    else:
                        tit = 3
                        grupos = grupos.order_by('grupo__tematica__carrera', 'matricula__inscripcion__persona__apellido1', 'matricula__inscripcion__persona__apellido2')
                        ws.write(encabezado, 2, 'CARRERA', stylebnotas)
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename =NOMINA DE ESTUDIANTES ' + str(carrera) + '.xls'
                    ws.col(0).width = 900
                    ws.col(1).width = 6000
                    ws.col(2).width = 5500
                    ws.col(tit).width = 5000
                    ws.col(tit + 1).width = 5000
                    ws.col(tit + 2).width = 5000
                    ws.col(tit + 3).width = 5000
                    if grupos.exists():
                        ws.write_merge(1, 1, 0, col1, u'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                        ws.write_merge(2, 2, 0, col1, u'VICERRECTORADO ACADÉMICO Y DE INVESTIGACIÓN', title)
                        ws.write_merge(3, 3, 0, col1, u'GESTIÓN TÉCNICA ACADÉMICA', title)
                        ws.write_merge(4, 4, 0, col1, u'PROCESO DE  TITULACIÓN', title)
                        ws.write_merge(5, 5, 0, col1, u'NÓMINA DE ESTUDIANTES APTOS PARA PROPUESTA PRÁCTICA', subtitle)
                        ws.write_merge(10, 10, 0, 1, u'ALTERNATIVA:', stylebnombre)
                        ws.write_merge(10, 10, 2, col1, u'EXAMEN DE GRADO O DE FIN DE CARRERA (DE CARÁCTER COMPLEXIVO)', stylebnombre)
                        ws.write_merge(11, 12, 0, 1, u'PROCESO DE TITULACIÓN \nGESTIÓN TÉCNICA ACADÉMICA:', stylebnombre)
                        ws.write_merge(11, 12, 2, col1, u'ING. VIVIANA GAIBOR HINOSTROZA, MSC.', stylebnombre)
                        ws.col(tit + 4).width = 2500
                        ws.write(6, 0, 'INICIO: ' + str(periodo.fechainicio) + ' FIN: ' + str(periodo.fechafin), subtitle)
                        ws.merge(6, 6, 0, col1)
                        tiem = datetime.today().date()
                        ws.write(0, 0, 'Milagro, ' + str(tiem), normalizsinborde)
                        ws.merge(0, 0, 0, col1)
                        ws.write(encabezado, 0, 'Nº', stylebnotas)
                        ws.write(encabezado, 1, 'APELLIDOS Y NOMBRES', stylebnotas)
                        ws.write(encabezado, tit, 'LÍNEA DE INVESTIGACIÓN', stylebnotas)
                        ws.write(encabezado, tit + 1, 'TEMA/VARIABLE', stylebnotas)
                        ws.write(encabezado, tit + 2, 'ACOMPAÑANTE', stylebnotas)
                        ws.write(encabezado, tit + 3, 'TRIBUNAL', stylebnotas)
                        ws.write(encabezado, tit + 4, 'ESTADO', stylebnotas)
                        normal.borders = borders
                        i = 1
                        for grupo in grupos:
                            integrante = ""
                            est = ''
                            if grupo.grupo.complexivodetallegrupo_set.filter(status=True, matricula__alternativa__grupotitulacion__periodogrupo=periodo).exclude(matricula__estado=8):
                                for com in grupo.grupo.complexivodetallegrupo_set.filter(status=True, matricula__alternativa__grupotitulacion__periodogrupo=periodo).exclude(matricula__estado=8):
                                    est = ''
                                    if com.matricula.reprobo_examen_complexivo():
                                        est = 'REPROBADO'
                                    integrante = integrante + u"• %s" % com.matricula + ' - [' + str(com.matricula.alternativa.paralelo) + '] ' + est + '\n'
                                ws.write(fil, 0, str(i), normal)
                                ws.write(fil, 1, u"%s" % integrante, normaliz)
                                if carrera == 0:
                                    ws.col(2).width = 4900
                                    ws.write(fil, 2, u"%s" % grupo.matricula.inscripcion.carrera, normal)
                                ws.write(fil, tit, u"%s" % grupo.grupo.tematica.tematica, normal)
                                subtema = " "
                                if grupo.grupo.subtema:
                                    subtema = grupo.grupo.subtema
                                ws.write(fil, tit + 1, u"%s" % subtema, normal)
                                ws.write(fil, tit + 2, u"%s" % grupo.grupo.tematica.tutor, normaliz)

                                ws.write(fil, tit + 4, u"%s" % (grupo.grupo.estado_propuesta().get_estado_display() if grupo.grupo.estado_propuesta() else ""), normaliz)
                                ws.write(fil, tit + 5, grupo.id, normaliz)
                                if grupo.grupo.presidentepropuesta:
                                    ws.write(fil, tit + 3, u"•%s\n•%s\n•%s" % (grupo.grupo.presidentepropuesta, grupo.grupo.secretariopropuesta, grupo.grupo.delegadopropuesta), normaliz)
                                else:
                                    ws.write(fil, tit + 3, u'', normaliz)

                                fil = fil + 1
                                i = i + 1
                        col1 = 3
                        col2 = 4
                        if not carrera == 0:
                            ws.write_merge(fil + 3, fil + 3, 1, 2, u'_____________________________', normalsinborde)
                            ws.write(fil + 4, 1, u"%s" % director, normalsinborde)
                            ws.merge(fil + 4, fil + 4, 1, 2, normalsinborde)
                            ws.write(fil + 5, 1, u'DIRECTOR(A) DE CARRERA', normalsinborde)
                            ws.merge(fil + 5, fil + 5, 1, 2, normalsinborde)
                            col1 = 4
                            col2 = 5
                        ws.write(fil + 3, col1, u'_____________________________\nING. VIVIANA GAIBOR,MSC.\nPROCESO DE TITULACIÓN', normalsinborde)
                        ws.merge(fil + 3, fil + 5, col1, col2, normalsinborde)

                    wb.save(response)
                    return response
                except Exception as es:
                    pass

            elif action == 'nominaestudiantes_numerohoras':
                try:
                    __author__ = 'Unemi'
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('reporte 1')
                    encabezado = 0
                    fil = encabezado + 1
                    col1 = 6
                    tit = 0
                    colaux = 1
                    carrera = 0
                    if 'car' in request.GET:
                        carrera = Carrera.objects.get(pk=request.GET['car'])
                    periodo = PeriodoGrupoTitulacion.objects.get(pk=request.GET['per'])
                    lstid = ComplexivoDetalleGrupo.objects.values_list('id').distinct('grupo').filter(status=True, grupo__tematica__periodo=periodo)

                    grupos = ComplexivoDetalleGrupo.objects.filter(pk__in=lstid, grupo__activo=True)
                    # para traer el total de tutorías
                    grupos_id = grupos.values_list('grupo_id', flat=True)
                    tutorias = ComplexivoAcompanamiento.objects.filter(grupo_id__in=grupos_id, status=True).values_list('grupo_id', flat=True).annotate(totcant=Count('grupo_id')).values('totcant').order_by('-totcant')
                    if tutorias:
                        colaux = tutorias[0]['totcant']
                    # para traer el total de revisiones
                    revisiones = ComplexivoPropuestaPractica.objects.filter(grupo_id__in=grupos_id, status=True).values_list('grupo_id', flat=True).annotate(totcant=Count('grupo_id')).values('totcant').order_by('-totcant')

                    if revisiones:
                        totrevisiones = revisiones[0]['totcant']
                    else:
                        totrevisiones = 0

                    if not carrera == 0:
                        tit = 4
                        col1 = 5
                        grupos = grupos.filter(grupo__tematica__carrera=carrera, grupo__tematica__periodo=periodo).order_by('matricula__inscripcion__persona__apellido1', 'matricula__inscripcion__persona__apellido2')
                        if grupos.exists():
                            ws.write_merge(7, 7, 0, 1, u'FACULTAD:', stylebnombre)
                            ws.write_merge(8, 8, 0, 1, u'CARRERA:', stylebnombre)
                            ws.write_merge(9, 9, 0, 1, u'DIRECTOR DE CARRERA:', stylebnombre)
                            director = u"%s" % (grupos[0].grupo.tematica.director)
                            ws.write(7, 2, u"%s" % carrera.coordinaciones().get().nombre, stylebnombre)
                            ws.merge(7, 7, 2, 6)
                            ws.write(8, 2, str(carrera), stylebnombre)
                            ws.merge(8, 8, 2, 6)
                            ws.write(9, 2, u"%s" % director, stylebnombre)
                            ws.merge(9, 9, 2, 6)
                    else:
                        tit = 5
                        grupos = grupos.order_by('grupo__tematica__carrera', 'matricula__inscripcion__persona__apellido1', 'matricula__inscripcion__persona__apellido2')
                        ws.col(3).width = 10000
                        ws.col(4).width = 10000
                        ws.write(encabezado, 3, 'CARRERA', stylebnotas)
                        ws.write(encabezado, 4, 'FACULTAD', stylebnotas)
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename =REPORTE MONITOREAR TUTORIA ' + str(carrera) + '.xls'
                    ws.col(0).width = 900
                    ws.col(1).width = 10000
                    ws.col(2).width = 14000
                    ws.col(tit).width = 10000
                    ws.col(tit + 1).width = 10000
                    ws.col(tit + 2).width = 10000
                    ws.col(tit + 3).width = 10000
                    ws.col(tit + 4).width = 10000
                    ws.col(tit + 5).width = 10000
                    ws.col(tit + 6).width = 10000
                    ws.col(tit + 7).width = 10000
                    ws.col(tit + 8).width = 10000
                    ws.col(tit + 9).width = 10000
                    ws.col(tit + 10).width = 10000
                    ws.col(tit + 11).width = 10000
                    ws.col(tit + 12).width = 10000
                    ws.col(tit + 13).width = 10000
                    ws.col(tit + 14).width = 10000
                    ws.col(tit + 15).width = 10000
                    ws.col(tit + 16).width = 10000
                    ws.col(tit + 17).width = 10000
                    ws.col(tit + 18).width = 10000
                    ws.col(tit + 19).width = 10000
                    ws.col(tit + 20).width = 10000
                    ws.col(tit + 21).width = 10000
                    if grupos.exists():

                        ws.write(encabezado, 0, 'Nº', stylebnotas)
                        ws.write(encabezado, 1, 'PERÍODO DE TITULACIÓN', stylebnotas)
                        ws.write(encabezado, 2, 'INTEGRANTES', stylebnotas)
                        ws.write(encabezado, tit, 'N. INTEGRANTES', stylebnotas)
                        ws.write(encabezado, tit + 1, 'ALTERNATIVA TITULACIÓN', stylebnotas)
                        ws.write(encabezado, tit + 2, 'PROYECTO DE INVESTIGACIÓN Ó LÍNEA DE INVESTIGACIÓN', stylebnotas)
                        ws.write(encabezado, tit + 3, 'TEMA/VARIABLE', stylebnotas)
                        ws.write(encabezado, tit + 4, 'DOCENTE TUTOR', stylebnotas)
                        ws.write(encabezado, tit + 5, 'EMAIL TUTOR', stylebnotas)
                        ws.write(encabezado, tit + 6, 'NÚMERO HORAS TUTORIAS', stylebnotas)
                        ws.write(encabezado, tit + 7, 'ESTADO TUTOR', stylebnotas)
                        coaux = 8
                        contut = 0
                        trev = 0
                        while coaux < colaux + 8:
                            contut += 1
                            ws.write(encabezado, tit + coaux, 'TUTORIA %s' % (contut), stylebnotas)
                            coaux += 1
                        ws.write(encabezado, tit + coaux, 'TRABAJO DE TITULACIÓN', stylebnotas)
                        ws.write(encabezado, tit + coaux + 1, 'FECHA ENTREGA TRABAJO DE TITULACIÓN', stylebnotas)
                        coaux += 2
                        while trev < totrevisiones:
                            trev += 1
                            ws.write(encabezado, tit + coaux, 'ESTADO DE REVISIÓN %s' % trev, stylebnotas)
                            coaux += 1
                            ws.write(encabezado, tit + coaux, 'OBSERVACIÓN DE REVISIÓN %s' % trev, stylebnotas)
                            coaux += 1

                        ws.write(encabezado, tit + coaux, 'ESTADO FINAL PREVIO SUSTENTACIÓN', stylebnotas)
                        ws.write(encabezado, tit + coaux + 1, 'TOTAL SUSTENTANTE', stylebnotas)
                        ws.write(encabezado, tit + coaux + 2, 'FECHA SUSTENTACIÓN', stylebnotas)
                        ws.write(encabezado, tit + coaux + 3, 'HORA SUSTENTACIÓN', stylebnotas)
                        ws.write(encabezado, tit + coaux + 4, 'LUGAR SUSTENTACIÓN', stylebnotas)

                        ws.write(encabezado, tit + coaux + 5, 'SUSTENTANTES', stylebnotas)

                        ws.write(encabezado, tit + coaux + 6, 'PRESIDENTE(A) SUSTENTACIÓN', stylebnotas)
                        ws.write(encabezado, tit + coaux + 7, 'SECRETARIO(A) SUSTENTACIÓN', stylebnotas)
                        ws.write(encabezado, tit + coaux + 8, 'INTEGRANTE SUSTENTACIÓN', stylebnotas)

                        ws.write(encabezado, tit + coaux + 9, 'ESTADO SUSTENTACIÓN DE GRUPOS', stylebnotas)
                        ws.write(encabezado, tit + coaux + 10, 'CANTIDAD CONVOCAR', stylebnotas)
                        ws.write(encabezado, tit + coaux + 11, 'CANTIDAD NO CONVOCAR', stylebnotas)
                        ws.write(encabezado, tit + coaux + 12, 'NOMBRE ESTUDIANTES CONVOCADOS', stylebnotas)
                        ws.write(encabezado, tit + coaux + 13, 'ESTUDIANTES APROBADOS', stylebnotas)
                        ws.write(encabezado, tit + coaux + 14, 'SUBEN ARCHIVO CORREGIDO', stylebnotas)
                        ws.write(encabezado, tit + coaux + 15, 'ESTADO ARCHIVO FINAL', stylebnotas)
                        normal.borders = borders
                        i = 1
                        coaux += 15
                        for grupo in grupos:
                            linea = ""
                            integrante = ""
                            sustentantes = ""
                            nomalternativa = ""
                            tutorias = ""
                            cadena = ""
                            totintegrantes = 0
                            numeroestado = 0
                            cuentafinalizados = 0
                            cuentavecesactivos = 0
                            cuentavecesnoaptos = 0
                            cuentavecesaptos = 0
                            cuentavecesreprobados = 0
                            unonoapto = 0
                            integrantesaprobados = 0
                            estudiantesaprobados = ''
                            estudiantesconvocadosaptos = ''
                            estudiantesconvocados = ""
                            if grupo.grupo.complexivodetallegrupo_set.filter(status=True, matricula__alternativa__grupotitulacion__periodogrupo=periodo).exclude(estado=3):

                                for com in grupo.grupo.complexivodetallegrupo_set.filter(status=True, matricula__alternativa__grupotitulacion__periodogrupo=periodo):
                                    est = ''
                                    estagrupo = ''
                                    if com.matricula.reprobo_examen_complexivo():
                                        est = '(REPROBADO)'
                                    if com.matricula.aprobo_examen_complexivo():
                                        est = '(APROBADO)'
                                        if com.estado == 1:
                                            estagrupo = '(ACTIVO)'
                                            numeroestado += 1
                                            sustentantes = sustentantes + u"• %s" % com.matricula + '\n'
                                    if com.matricula.encurso_examen_complexivo():
                                        est = '(EN CURSO)'
                                        if com.estado == 1:
                                            estagrupo = '(ACTIVO)'
                                            numeroestado += 1
                                    if com.estado == 3:
                                        estagrupo = '(INCONSISTENCIA TUTORIA)'
                                    apto = 'PENDIENTE'
                                    if com.matricula.cumplerequisitos == 2:
                                        apto = 'APTO'
                                        cuentavecesaptos += 1
                                        estudiantesconvocados = estudiantesconvocados + u"• %s" % com.matricula + '-[' + str(com.matricula.alternativa.paralelo) + '] ' + '\n'
                                    if com.matricula.cumplerequisitos == 3:
                                        apto = 'NO APTO'
                                        unonoapto = 1
                                        cuentavecesnoaptos += 1
                                    integrante = integrante + u"• %s" % com.matricula + '-[' + str(com.matricula.alternativa.paralelo) + '] ' + est + '-' + estagrupo + '-' + str(apto) + '\n'
                                    nomalternativa = com.matricula.alternativa.tipotitulacion
                                    totintegrantes += 1
                                    verestadosustentar = 'REPROBADO' if com.matricula.reprobo_examen_complexivo() else com.matricula.get_estado_display()

                                    if verestadosustentar == 'APROBADO':
                                        estudiantesaprobados = estudiantesaprobados + u"• %s" % com.matricula + '-[' + str(com.matricula.alternativa.paralelo) + '] ' + ','
                                        integrantesaprobados += 1
                                        cuentafinalizados = 1
                                    if verestadosustentar == 'ACTIVO':
                                        cuentavecesactivos += 1
                                    if verestadosustentar == 'REPROBADO':
                                        cuentavecesreprobados += 1

                                cantidadconvocar = 0
                                cantidadnoconvocar = 0
                                estadosustentar = ''

                                if cuentafinalizados == 1:
                                    estadosustentar = 'FINALIZADO'
                                    cantidadconvocar = integrantesaprobados
                                    cantidadnoconvocar = cuentavecesreprobados
                                else:
                                    if totintegrantes == cuentavecesactivos and totintegrantes == cuentavecesnoaptos:
                                        estadosustentar = 'NO CONVOCAR'
                                        cantidadnoconvocar = cuentavecesnoaptos

                                    else:
                                        if totintegrantes == cuentavecesactivos and totintegrantes == cuentavecesaptos:
                                            estadosustentar = 'CONVOCAR'
                                            cantidadconvocar = cuentavecesaptos
                                            cantidadnoconvocar = cuentavecesnoaptos
                                            estudiantesconvocadosaptos = estudiantesconvocados
                                        else:
                                            if totintegrantes == cuentavecesactivos and unonoapto == 1:
                                                estadosustentar = 'REVISAR CONVOCATORIA'
                                                cantidadconvocar = cuentavecesaptos
                                                cantidadnoconvocar = cuentavecesnoaptos
                                                estudiantesconvocadosaptos = estudiantesconvocados
                                            else:
                                                if totintegrantes == cuentavecesreprobados:
                                                    estadosustentar = 'NUEVO PROCESO'
                                                    cantidadnoconvocar = cuentavecesreprobados
                                subirarchivofinalgrupo = 'NO'
                                archivofinalgruposubido = 'NO'
                                if grupo.grupo.subirarchivofinalgrupo:
                                    subirarchivofinalgrupo = 'SI'
                                    if grupo.grupo.archivofinalgrupo:
                                        archivofinalgruposubido = 'SI'

                                if grupo.grupo.archivofinalgrupo:
                                    estadogrupo = grupo.grupo.get_estadoarchivofinalgrupo_display()
                                else:
                                    estadogrupo = 'PENDIENTE'

                                ws.write(fil, 0, str(i), normal)
                                ws.write(fil, 1, u"%s" % grupo.grupo.tematica.periodo.nombre, normaliz)
                                ws.write(fil, 2, u"%s" % integrante, normaliz)
                                if carrera == 0:
                                    ws.col(2).width = 4900
                                    ws.write(fil, 3, u"%s" % grupo.matricula.inscripcion.carrera, normal)
                                    ws.write(fil, 4, u"%s" % grupo.matricula.inscripcion.coordinacion, normal)

                                ws.write(fil, tit, u"%s" % totintegrantes, normal)
                                ws.write(fil, tit + 1, u"%s" % nomalternativa, normal)
                                ws.write(fil, tit + 2, u"%s" % grupo.grupo.tematica.tematica, normal)
                                subtema = " "
                                if grupo.grupo.subtema:
                                    subtema = grupo.grupo.subtema
                                ws.write(fil, tit + 3, u"%s" % subtema, normal)
                                ws.write(fil, tit + 4, u"%s" % (str(grupo.grupo.tematica.tutor)), normaliz)
                                ws.write(fil, tit + 7, u"%s" % ("INACTIVO" if not grupo.grupo.activo else "ACTIVO"), normaliz)
                                ws.write(fil, tit + 5, u"%s" % grupo.grupo.tematica.tutor.participante.persona.emailinst, normaliz)
                                ws.write(fil, tit + 6, u"%s" % str(grupo.grupo.horas_totales_tutorias_grupo()), normal)
                                colum = 8

                                tutorias = grupo.grupo.complexivoacompanamiento_set.filter(status=True)
                                for tut in tutorias:
                                    cadena = str(tut.fecha) + '-' + str(tut.horainicio) + '-' + str(tut.horafin) + '-' + str(tut.observaciones)
                                    ws.write(fil, tit + colum, u"%s" % cadena, normal)
                                    colum = colum + 1
                                archivo = 'No'
                                fecha = ''
                                col = 8 + colaux
                                caux2 = 0
                                if ComplexivoPropuestaPractica.objects.filter(grupo=grupo.grupo, status=True).exists():
                                    for pro in ComplexivoPropuestaPractica.objects.filter(grupo=grupo.grupo, status=True):
                                        texto = ''
                                        if pro.complexivopropuestapracticaarchivo_set.filter(tipo=1, status=True).exists():
                                            fecha = pro.complexivopropuestapracticaarchivo_set.filter(tipo=1, status=True)[0].fecha.strftime('%Y-%m-%d')
                                            archivo = 'Si'
                                        texto = str(pro.fecharevision) + '-' + str(pro.observacion)
                                        ws.write(fil, tit + col, u"%s" % str(pro.get_estado_display()), normal)
                                        col = col + 1
                                        ws.write(fil, tit + col, u"%s" % texto, normal)
                                        col = col + 1
                                        caux2 += 2

                                # col = ((coaux - col)+8)+((totrevisiones*2)+(caux2))
                                col = coaux - 15
                                print(col)

                                ws.write(fil, tit + col, u"%s" % (grupo.grupo.estado_propuesta().get_estado_display() if grupo.grupo.estado_propuesta() else ""), normal)
                                ws.write(fil, tit + col + 1, u"%s" % numeroestado, normal)
                                ws.write(fil, tit + col + 2, u'%s' % (grupo.grupo.fechadefensa if grupo.grupo.fechadefensa else ''), normal)
                                ws.write(fil, tit + col + 3, u'%s' % (grupo.grupo.horadefensa if grupo.grupo.horadefensa else ''), normal)
                                ws.write(fil, tit + col + 4, u'%s' % (grupo.grupo.lugardefensa if grupo.grupo.lugardefensa else ''), normal)

                                ws.write(fil, tit + col + 5, u'%s' % sustentantes, normal)

                                ws.write(fil, tit + col + 6, u'%s' % (grupo.grupo.presidentepropuesta if grupo.grupo.presidentepropuesta else ''), normal)
                                ws.write(fil, tit + col + 7, u'%s' % (grupo.grupo.secretariopropuesta if grupo.grupo.secretariopropuesta else ''), normal)
                                ws.write(fil, tit + col + 8, u'%s' % (grupo.grupo.delegadopropuesta if grupo.grupo.delegadopropuesta else ''), normal)

                                ws.write(fil, tit + col + 9, u'%s' % estadosustentar, normal)
                                ws.write(fil, tit + col + 10, u'%s' % cantidadconvocar, normal)
                                ws.write(fil, tit + col + 11, u'%s' % cantidadnoconvocar, normal)
                                ws.write(fil, tit + col + 12, u'%s' % estudiantesconvocadosaptos, normal)
                                ws.write(fil, tit + col + 13, u'%s' % estudiantesaprobados, normal)
                                ws.write(fil, tit + col + 14, u'%s' % subirarchivofinalgrupo, normal)
                                ws.write(fil, tit + col + 15, u'%s' % estadogrupo, normal)
                                fil = fil + 1
                                i = i + 1
                            col1 = 3
                            col2 = 4
                            if not carrera == 0:
                                ws.write_merge(fil + 3, fil + 3, 1, 2, u'_____________________________', normalsinborde)
                                ws.write(fil + 4, 1, u"%s" % director, normalsinborde)
                                ws.merge(fil + 4, fil + 4, 1, 2, normalsinborde)
                                ws.write(fil + 5, 1, u'DIRECTOR(A) DE CARRERA', normalsinborde)
                                ws.merge(fil + 5, fil + 5, 1, 2, normalsinborde)
                                col1 = 4
                                col2 = 5

                        wb.save(response)
                        return response
                except Exception as es:
                    print('Error on line {} - {}'.format(sys.exc_info()[-1].tb_lineno, es))

                    pass

            elif action == 'reporte_revision':
                try:
                    __author__ = 'Unemi'
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('reporte 1')
                    encabezado = 0
                    fil = encabezado + 1
                    director = ''
                    col1 = 6
                    tit = 5
                    carrera = 0
                    idper = int(request.GET['idper'])
                    codifacu = int(request.GET['codifacu'])
                    codicarr = int(request.GET['codicarr'])

                    periodo = PeriodoGrupoTitulacion.objects.get(pk=request.GET['idper'])
                    lstid = ComplexivoDetalleGrupo.objects.values_list('id').distinct('grupo').filter(status=True,
                                                                                                      grupo__tematica__periodo=periodo)
                    grupos = ComplexivoDetalleGrupo.objects.filter(pk__in=lstid, grupo__activo=True)

                    if idper > 0 and codifacu == 0 and codicarr == 0:
                        grupos = grupos.order_by('grupo__tematica__carrera', 'matricula__inscripcion__persona__apellido1', 'matricula__inscripcion__persona__apellido2')
                    if codifacu > 0 and codicarr == 0:
                        grupos = grupos.filter(grupo__tematica__carrera__coordinacion__id=codifacu, grupo__tematica__periodo=periodo).order_by('matricula__inscripcion__persona__apellido1', 'matricula__inscripcion__persona__apellido2')

                    if codicarr > 0:
                        grupos = grupos.filter(grupo__tematica__carrera__id=codicarr, grupo__tematica__periodo=periodo).order_by('matricula__inscripcion__persona__apellido1', 'matricula__inscripcion__persona__apellido2')

                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename =REPORTE REVISION TEMA ' + str(carrera) + '.xls'
                    ws.col(0).width = 900
                    ws.col(1).width = 6000
                    ws.col(2).width = 14000
                    ws.col(tit).width = 10000
                    ws.col(tit + 1).width = 10000
                    ws.col(tit + 2).width = 10000
                    ws.col(tit + 3).width = 10000
                    ws.col(tit + 4).width = 10000
                    ws.col(tit + 5).width = 10000
                    ws.col(tit + 6).width = 10000
                    ws.col(tit + 7).width = 10000
                    ws.col(tit + 8).width = 10000

                    tit = 5
                    if grupos.exists():

                        ws.write(encabezado, 0, 'Nº', stylebnotas)
                        ws.write(encabezado, 1, 'PERÍODO DE TITULACIÓN', stylebnotas)
                        ws.write(encabezado, 2, 'INTEGRANTES', stylebnotas)
                        ws.write(encabezado, 3, 'CARRERA', stylebnotas)
                        ws.write(encabezado, 4, 'FACULTAD', stylebnotas)
                        ws.write(encabezado, tit, 'N. INTEGRANTES', stylebnotas)
                        ws.write(encabezado, tit + 1, 'ALTERNATIVA DE TITULACIÓN', stylebnotas)
                        ws.write(encabezado, tit + 2, 'PROYECTO DE INVESTIGACIÓN O LÍNEA DE INVESTIGACIÓN', stylebnotas)
                        ws.write(encabezado, tit + 3, 'TEMA/VARIABLE', stylebnotas)
                        ws.write(encabezado, tit + 4, 'DOCENTE TUTOR', stylebnotas)
                        ws.write(encabezado, tit + 5, 'EMAIL TUTOR', stylebnotas)
                        ws.write(encabezado, tit + 6, 'NÚMERO DE HORAS TUTORIAS', stylebnotas)
                        ws.write(encabezado, tit + 7, 'ESTADO REVISIÓN TEMA', stylebnotas)
                        ws.write(encabezado, tit + 8, 'OBSERVACIONES', stylebnotas)

                        normal.borders = borders
                        i = 1
                        for grupo in grupos:
                            integrante = ""
                            sustentantes = ""
                            nomalternativa = ""

                            totintegrantes = 0
                            numeroestado = 0
                            cuentafinalizados = 0
                            cuentavecesactivos = 0
                            cuentavecesnoaptos = 0
                            cuentavecesaptos = 0
                            cuentavecesreprobados = 0
                            unonoapto = 0
                            integrantesaprobados = 0
                            estudiantesaprobados = ''
                            estudiantesconvocadosaptos = ''
                            estudiantesconvocados = ""
                            if grupo.grupo.complexivodetallegrupo_set.filter(status=True, matricula__alternativa__grupotitulacion__periodogrupo=periodo).exclude(estado=3):

                                for com in grupo.grupo.complexivodetallegrupo_set.filter(status=True, matricula__alternativa__grupotitulacion__periodogrupo=periodo):
                                    est = ''
                                    estagrupo = ''
                                    if com.matricula.reprobo_examen_complexivo():
                                        est = '(REPROBADO)'
                                    if com.matricula.aprobo_examen_complexivo():
                                        est = '(APROBADO)'
                                        if com.estado == 1:
                                            estagrupo = '(ACTIVO)'
                                            numeroestado += 1
                                            sustentantes = sustentantes + u"• %s" % com.matricula + '\n'
                                    if com.matricula.encurso_examen_complexivo():
                                        est = '(EN CURSO)'
                                        if com.estado == 1:
                                            estagrupo = '(ACTIVO)'
                                            numeroestado += 1
                                    if com.estado == 3:
                                        estagrupo = '(INCONSISTENCIA TUTORIA)'
                                    apto = 'PENDIENTE'
                                    if com.matricula.cumplerequisitos == 2:
                                        apto = 'APTO'
                                        cuentavecesaptos += 1
                                        estudiantesconvocados = estudiantesconvocados + u"• %s" % com.matricula + '-[' + str(com.matricula.alternativa.paralelo) + '] ' + '\n'
                                    if com.matricula.cumplerequisitos == 3:
                                        apto = 'NO APTO'
                                        unonoapto = 1
                                        cuentavecesnoaptos += 1
                                    integrante = integrante + u"• %s" % com.matricula + '-[' + str(com.matricula.alternativa.paralelo) + '] ' + est + '-' + estagrupo + '-' + str(apto) + '\n'
                                    nomalternativa = com.matricula.alternativa.tipotitulacion
                                    totintegrantes += 1
                                    verestadosustentar = 'REPROBADO' if com.matricula.reprobo_examen_complexivo() else com.matricula.get_estado_display()

                                    if verestadosustentar == 'APROBADO':
                                        estudiantesaprobados = estudiantesaprobados + u"• %s" % com.matricula + '-[' + str(com.matricula.alternativa.paralelo) + '] ' + ','
                                        integrantesaprobados += 1
                                    if verestadosustentar == 'ACTIVO':
                                        cuentavecesactivos += 1
                                    if verestadosustentar == 'REPROBADO':
                                        cuentavecesreprobados += 1

                                ws.write(fil, 0, str(i), normal)
                                ws.write(fil, 1, u"%s" % grupo.grupo.tematica.periodo, normaliz)
                                ws.write(fil, 2, u"%s" % integrante, normaliz)
                                if carrera == 0:
                                    ws.col(3).width = 4900
                                    ws.col(4).width = 4900
                                    ws.write(fil, 3, u"%s" % grupo.matricula.inscripcion.carrera, normal)
                                    ws.write(fil, 4, u"%s" % grupo.matricula.inscripcion.coordinacion, normal)

                                ws.write(fil, tit, u"%s" % totintegrantes, normal)
                                ws.write(fil, tit + 1, u"%s" % nomalternativa, normal)
                                ws.write(fil, tit + 2, u"%s" % grupo.grupo.tematica.tematica, normal)
                                subtema = " "
                                if grupo.grupo.subtema:
                                    subtema = grupo.grupo.subtema

                                ws.write(fil, tit + 3, u"%s" % subtema, normal)
                                ws.write(fil, tit + 4, u"%s" % (str(grupo.grupo.tematica.tutor)), normaliz)
                                ws.write(fil, tit + 5, u"%s" % grupo.grupo.tematica.tutor.participante.persona.emailinst, normaliz)
                                ws.write(fil, tit + 6, u"%s" % str(grupo.grupo.horas_totales_tutorias_grupo()), normal)

                                fil = fil + 1
                                i = i + 1

                            if not carrera == 0:
                                ws.write_merge(fil + 3, fil + 3, 1, 2, u'_____________________________', normalsinborde)
                                ws.write(fil + 4, 1, u"%s" % director, normalsinborde)
                                ws.merge(fil + 4, fil + 4, 1, 2, normalsinborde)
                                ws.write(fil + 5, 1, u'DIRECTOR(A) DE CARRERA', normalsinborde)
                                ws.merge(fil + 5, fil + 5, 1, 2, normalsinborde)

                        wb.save(response)
                        return response
                except Exception as es:
                    pass

            elif action == 'listadoalumnosexcell':
                try:
                    __author__ = 'Unemi'
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('reporte 1')
                    fuentenormal = easyxf('font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
                    encabezado = 0
                    fil = encabezado + 1
                    director = ''
                    col1 = 6
                    tit = 5
                    carrera = 0
                    materia = Materia.objects.get(pk=int(encrypt(request.GET['idmat'])))

                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename =LISTADO_' + str(materia.asignaturamalla.asignatura.nombre[0:30]) + '_' + str(materia.paralelo) + '.xls'
                    row_num = 0
                    columns = [
                        (u"#N", 1500),
                        (u"CARRERA", 10000),
                        (u"MATERIA", 10000),
                        (u"PARALELO", 4000),
                        (u"CEDULA", 4000),
                        (u"NOMBRES COMPLETOS", 10000),
                        (u"CORREO PERSONAL", 8000),
                        (u"CORREO UNEMI", 8000),
                        (u"TELÉFONO", 5000),
                        (u"CELULAR", 5000),
                        (u"CUMPLE REQUISITOS", 4000),
                        (u"NUMERO MEMORANDO", 10000)
                    ]
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], fuentenormal)
                        ws.col(col_num).width = columns[col_num][1]

                    listadorequisitos = RequisitoMateriaUnidadIntegracionCurricular.objects.filter(materia=materia, titulacion=True, activo=True, status=True)
                    for inscrito in MateriaTitulacion.objects.filter(materiaasignada__materia=materia, status=True).order_by('materiaasignada__matricula__inscripcion__persona__apellido1', 'materiaasignada__matricula__inscripcion__persona__apellido2', 'materiaasignada__matricula__inscripcion__persona__nombres'):
                        row_num += 1
                        cumplerequisitos = 'SI'
                        alumno = inscrito.materiaasignada.matricula.inscripcion.persona
                        for lis in listadorequisitos:
                            valida = lis.run(inscrito.materiaasignada.matricula.inscripcion.id)
                            if not valida:
                                cumplerequisitos = 'NO'
                        ws.write(row_num, 0, row_num, fuentenormal)
                        ws.write(row_num, 1, inscrito.materiaasignada.matricula.inscripcion.carrera.nombre, fuentenormal)
                        ws.write(row_num, 2, materia.asignaturamalla.asignatura.nombre, fuentenormal)
                        ws.write(row_num, 3, materia.paralelo, fuentenormal)
                        ws.write(row_num, 4, alumno.identificacion(), fuentenormal)
                        ws.write(row_num, 5, alumno.nombre_completo_inverso(), fuentenormal)
                        ws.write(row_num, 6, alumno.email, fuentenormal)
                        ws.write(row_num, 7, alumno.emailinst, fuentenormal)
                        ws.write(row_num, 8, alumno.telefono_conv, fuentenormal)
                        ws.write(row_num, 9, alumno.telefono, fuentenormal)
                        ws.write(row_num, 10, cumplerequisitos, fuentenormal)
                        ws.write(row_num, 11, inscrito.numeromemo, fuentenormal)
                    wb.save(response)
                    return response
                except Exception as es:
                    pass

            elif action == 'listadoalumnosexcell2':
                try:
                    __author__ = 'Unemi'
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('reporte 1')
                    fuentenormal = easyxf('font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
                    encabezado = 0
                    fil = encabezado + 1
                    director = ''
                    col1 = 6
                    tit = 5
                    carrera = 0
                    materia = Materia.objects.get(pk=int(encrypt(request.GET['idmat'])))

                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename =LISTADO_' + str(materia.asignaturamalla.asignatura.nombre[0:30]) + '_' + str(materia.paralelo) + '.xls'
                    row_num = 0
                    columns = [
                        (u"#N", 1500),
                        (u"FACULTAD", 10000),
                        (u"CARRERA", 10000),
                        (u"ALTERNATIVA TITULACIÓN", 10000),
                        (u"CEDULA", 4000),
                        (u"NOMBRES COMPLETOS", 10000),
                        (u"NÚMERO DE ACTA", 10000),
                        (u"NOTA FINAL", 10000),
                        (u"ESTADO FINAL", 10000),
                        (u"PERIODO ACADÉMICO", 10000),
                        (u"NUMERO MEMORANDO", 10000),
                        (u"EMAIL INSTITUCIONAL", 10000),
                        (u"EMAIL PERSONAL", 10000),
                        (u"TELEFONO CONVENCIONAL", 10000),
                        (u"TELEFONO CELULAR", 10000)
                    ]
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], fuentenormal)
                        ws.col(col_num).width = columns[col_num][1]

                    listadorequisitos = RequisitoMateriaUnidadIntegracionCurricular.objects.filter(materia=materia, titulacion=True, activo=True, status=True)
                    for inscrito in MateriaTitulacion.objects.filter(materiaasignada__materia=materia, materiaasignada__status=True, status=True).order_by('materiaasignada__matricula__inscripcion__persona__apellido1', 'materiaasignada__matricula__inscripcion__persona__apellido2', 'materiaasignada__matricula__inscripcion__persona__nombres'):
                        row_num += 1

                        alumno = inscrito.materiaasignada.matricula.inscripcion.persona
                        acta = 'SUS-%s-%s-%s' %(inscrito.materiaasignada.matricula.inscripcion.carrera.abrsustentacion,inscrito.numeroacta,inscrito.materiaasignada.materia.nivel.periodo.fin)
                        ws.write(row_num, 0, row_num, fuentenormal)
                        ws.write(row_num, 1, str(inscrito.materiaasignada.matricula.inscripcion.coordinacion), fuentenormal)
                        ws.write(row_num, 2, inscrito.materiaasignada.matricula.inscripcion.carrera.nombre, fuentenormal)
                        ws.write(row_num, 3, "EXAMEN COMPLEXIVO", fuentenormal)
                        ws.write(row_num, 4, alumno.identificacion(), fuentenormal)
                        ws.write(row_num, 5, alumno.nombre_completo_inverso(), fuentenormal)
                        ws.write(row_num, 6, acta, fuentenormal)
                        ws.write(row_num, 7, str(inscrito.materiaasignada.notafinal), fuentenormal)
                        ws.write(row_num, 8, str(inscrito.materiaasignada.estado), fuentenormal)
                        ws.write(row_num, 9, str(inscrito.materiaasignada.materia.nivel.periodo), fuentenormal)
                        ws.write(row_num, 10, inscrito.numeromemo, fuentenormal)
                        ws.write(row_num, 11, alumno.emailinst, fuentenormal)
                        ws.write(row_num, 12, alumno.email, fuentenormal)
                        ws.write(row_num, 13, alumno.telefono_conv, fuentenormal)
                        ws.write(row_num, 14, alumno.telefono, fuentenormal)
                    wb.save(response)
                    return response
                except Exception as es:
                    pass

            elif action == 'listadoalumnosexcellrubrica':
                try:
                    __author__ = 'Unemi'
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('reporte 1')
                    fuentenormal = easyxf('font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
                    encabezado = 0
                    fil = encabezado + 1
                    director = ''
                    col1 = 6
                    tit = 5
                    carrera = 0
                    materia = Materia.objects.get(pk=int(encrypt(request.GET['idmat'])))

                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename =LISTADO_' + str(materia.asignaturamalla.asignatura.nombre[0:30]) + '_' + str(materia.paralelo) + '.xls'
                    row_num = 0
                    columns = [
                        (u"#N", 1500),
                        (u"FACULTAD", 10000),
                        (u"CARRERA", 10000),
                        (u"ALTERNATIVA TITULACIÓN", 10000),
                        (u"CEDULA", 4000),
                        (u"NOMBRES COMPLETOS", 10000),
                        (u"NÚMERO DE ACTA", 10000),
                        (u"NOTA FINAL", 10000),
                        (u"ESTADO FINAL", 10000),
                        (u"PERIODO ACADÉMICO", 10000),
                        (u"NUMERO MEMORANDO", 10000),
                        (u"EMAIL INSTITUCIONAL", 10000),
                        (u"EMAIL PERSONAL", 10000),
                        (u"TELEFONO CONVENCIONAL", 10000),
                        (u"TELEFONO CELULAR", 10000)
                    ]
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], fuentenormal)
                        ws.col(col_num).width = columns[col_num][1]

                    listadorequisitos = RequisitoMateriaUnidadIntegracionCurricular.objects.filter(materia=materia, titulacion=True, activo=True, status=True)
                    for inscrito in MateriaTitulacion.objects.filter(materiaasignada__materia=materia, materiaasignada__status=True, status=True).order_by('materiaasignada__matricula__inscripcion__persona__apellido1', 'materiaasignada__matricula__inscripcion__persona__apellido2', 'materiaasignada__matricula__inscripcion__persona__nombres'):
                        row_num += 1

                        alumno = inscrito.materiaasignada.matricula.inscripcion.persona
                        acta = 'SUS-%s-%s-%s' %(inscrito.materiaasignada.matricula.inscripcion.carrera.abrsustentacion,inscrito.numeroacta,inscrito.materiaasignada.materia.nivel.periodo.fin)
                        ws.write(row_num, 0, row_num, fuentenormal)
                        ws.write(row_num, 1, str(inscrito.materiaasignada.matricula.inscripcion.coordinacion), fuentenormal)
                        ws.write(row_num, 2, inscrito.materiaasignada.matricula.inscripcion.carrera.nombre, fuentenormal)
                        ws.write(row_num, 3, "EXAMEN COMPLEXIVO", fuentenormal)
                        ws.write(row_num, 4, alumno.identificacion(), fuentenormal)
                        ws.write(row_num, 5, alumno.nombre_completo_inverso(), fuentenormal)
                        ws.write(row_num, 6, acta, fuentenormal)
                        ws.write(row_num, 7, str(inscrito.notafinal), fuentenormal)
                        ws.write(row_num, 8, str(inscrito.materiaasignada.estado), fuentenormal)
                        ws.write(row_num, 9, str(inscrito.materiaasignada.materia.nivel.periodo), fuentenormal)
                        ws.write(row_num, 10, inscrito.numeromemo, fuentenormal)
                        ws.write(row_num, 11, alumno.emailinst, fuentenormal)
                        ws.write(row_num, 12, alumno.email, fuentenormal)
                        ws.write(row_num, 13, alumno.telefono_conv, fuentenormal)
                        ws.write(row_num, 14, alumno.telefono, fuentenormal)
                    wb.save(response)
                    return response
                except Exception as es:
                    pass

            elif action == 'nominatutor':
                try:
                    carrera = None
                    if 'c' in request.GET:
                        carrera = Carrera.objects.get(pk=int(request.GET['c']))
                    periodo = PeriodoGrupoTitulacion.objects.get(pk=request.GET['p'])
                    if carrera:
                        tematicas = ComplexivoTematica.objects.filter(carrera=carrera, periodo=periodo, status=True).order_by('tutor__participante__persona__apellido1', 'tutor__participante__persona__apellido2').distinct()
                        grupos = ComplexivoDetalleGrupo.objects.filter(status=True, grupo__tematica__carrera=carrera, grupo__tematica__periodo=periodo)
                    else:
                        tematicas = ComplexivoTematica.objects.filter(periodo=periodo, status=True).order_by('tutor__participante__persona__apellido1', 'tutor__participante__persona__apellido2').distinct()
                        grupos = ComplexivoDetalleGrupo.objects.filter(status=True, grupo__tematica__periodo=periodo)
                    if tematicas.exists():
                        __author__ = 'Unemi'
                        wb = Workbook(encoding='utf-8')
                        ws = wb.add_sheet('reporte 1')
                        ws.col(0).width = 700
                        ws.col(1).width = 9000
                        ws.col(2).width = 9000
                        ws.col(3).width = 9000
                        ws.write(0, 0, 'Milagro, ' + str(datetime.today().date()), normalizsinborde)
                        ws.merge(0, 0, 0, 7)
                        ws.write_merge(1, 1, 0, 5, u'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                        ws.write_merge(2, 2, 0, 5, u'VICERRECTORADO ACADÉMICO Y DE INVESTIGACIÓN', title)
                        ws.write_merge(3, 3, 0, 5, u'GESTIÓN TÉCNICA ACADÉMICA', title)
                        ws.write_merge(4, 4, 0, 5, u'PROCESO DE  TITULACIÓN', title)
                        ws.write_merge(5, 5, 0, 5, u'NÓMINA DE ACOMPAÑANTES', subtitle)
                        if carrera:
                            ws.write_merge(6, 6, 0, 1, u'FACULTAD:', stylebnombre)
                            ws.write_merge(7, 7, 0, 1, u'CARRERA:', stylebnombre)
                        ws.write_merge(8, 8, 0, 1, u'ALTERNATIVA:', stylebnombre)
                        ws.write_merge(8, 8, 2, 5, u'EXAMEN DE GRADO O DE FIN DE CARRERA (DE CARÁCTER COMPLEXIVO)', stylebnombre)
                        if carrera:
                            ws.write_merge(9, 9, 0, 1, u'DIRECTOR DE CARRERA:', stylebnombre)
                        ws.write_merge(10, 11, 0, 1, u'PROCESO DE TITULACIÓN GESTIÓN TÉCNICA ACADÉMICA:', stylebnombre)
                        ws.write_merge(10, 11, 2, 5, u'ING. VIVIANA GAIBOR HINOSTROZA, MSC.', stylebnombre)
                        response = HttpResponse(content_type="application/ms-excel")
                        response['Content-Disposition'] = 'attachment; filename =NOMINA DE ACOMPAÑANTES ' + str(carrera) + '.xls'
                        director = ''
                        if carrera:
                            director = str(grupos[0].grupo.tematica.director)
                            ws.write(6, 2, str(carrera.coordinaciones().get().nombre), stylebnombre)
                            ws.merge(6, 6, 2, 3)
                            ws.write(7, 2, str(carrera), stylebnombre)
                            ws.merge(7, 7, 2, 3)
                            ws.write(9, 2, str(director), stylebnombre)
                            ws.merge(9, 9, 2, 3)

                        encabezado = 12
                        ws.write(encabezado, 0, 'Nº', stylebnotas)
                        ws.write(encabezado, 1, 'ACOMPAÑANTE', stylebnotas)
                        ws.write(encabezado, 2, 'LÍNEA DE INVESTIGACI', stylebnotas)
                        ws.write(encabezado, 3, 'ESTUDIANTES', stylebnotas)
                        ws.write(encabezado, 4, 'FACULTAD', stylebnotas)
                        ws.write(encabezado, 5, 'CARRERA', stylebnotas)
                        normal.borders = borders
                        fil = encabezado + 1
                        i = 1
                        b = 0
                        for tematica in tematicas:
                            c = 0
                            estudiante = ''
                            detalles = ComplexivoDetalleGrupo.objects.filter(grupo__tematica__id=tematica.id, status=True, matricula__alternativa__grupotitulacion__periodogrupo=periodo).exclude(matricula__estado=8)
                            if detalles:
                                for integrante in detalles:
                                    est = ''
                                    if integrante.matricula.reprobo_examen_complexivo():
                                        est = '(REPROBADO)'
                                    estudiante = estudiante + '\n•' + str(integrante.matricula) + ' ' + str(integrante.matricula.alternativa.paralelo) + ' ' + str(est)
                            if estudiante != '':
                                ws.write(fil, 0, str(i), normal)
                                ws.write(fil, 2, str(tematica.tematica), normal)
                                ws.write(fil, 3, str(estudiante), normaliz)
                                ws.write(fil, 4, str(tematica.carrera.coordinaciones()[0].nombre), normal)
                                ws.write(fil, 5, str(tematica.carrera.nombre), normal)
                                fil = fil + 1
                                b = 1
                                c = c + 1
                            if b == 1:
                                if tematica.tutor.participante.persona:
                                    tutor = tematica.tutor.participante.persona.nombre_completo_inverso()
                                else:
                                    tutor = str(tematica.tutor.participante.apellido) + ' ' + str(tematica.tutor.participante.nombre)
                                ws.write(fil - c, 1, u"" + str(tutor), normalcenter)
                                ws.merge(fil - c, (fil - 1), 1, 1, normal)
                                b = 0

                        # for tutor in tutores :
                        #     tu=tutor[1]+" "+tutor[2]+" "+tutor[3]
                        #     car = Carrera.objects.filter(id=tutor[4], status=True)[0]
                        #     ws.write(fil, 4, str(car.coordinaciones()[0].nombre), normal)
                        #     ws.write(fil, 5, str(car.nombre), normal)
                        #     c=0
                        #     if carrera:
                        #         tematicas = ComplexivoTematica.objects.filter(carrera=carrera, periodo=periodo, status=True,tutor__participante__persona__id=tutor[0])
                        #     else:
                        #         tematicas = ComplexivoTematica.objects.filter(periodo=periodo, status=True, tutor__participante__persona__id=tutor[0])
                        #     for tema in tematicas:
                        #         estu = ''
                        #         for grupo in grupos:
                        #             if grupo.grupo.tematica.tutor.participante.persona_id == tutor[0] and grupo.grupo.tematica==tema:
                        #                 est = ''
                        #                 if grupo.matricula.reprobo_examen_complexivo():
                        #                     est= 'R'
                        #                 estu=estu+'\n•'+str(grupo.matricula)+' '+str(grupo.matricula.alternativa.paralelo)+' '+str(est)
                        #         if estu!='':
                        #             ws.write(fil, 0, str(i), normal)
                        #             ws.write(fil, 2, str(tema.tematica), normal)
                        #             ws.write(fil,3,str(estu),normaliz)
                        #             fil = fil + 1
                        #             i = i + 1
                        #             b=1
                        #             c=c+1
                        #     if b==1:
                        #         ws.write(fil-c, 1, u""+tu, normalcenter)
                        #         ws.merge(fil-c,(fil-1),1,1,normal)
                        #         b=0

                        ws.write(fil + 3, 3, u"_____________________________\nING. VIVIANA GAIBOR,MSC.\nPROCESO DE TITULACIÓN", normalsinborde)
                        ws.merge(fil + 3, fil + 5, 3, 3, normalsinborde)
                        ws.write_merge(fil + 3, fil + 3, 0, 1, u'_____________________________', normalsinborde)
                        ws.write(fil + 4, 0, str(director), normalsinborde)
                        ws.merge(fil + 4, fil + 4, 0, 1, normalsinborde)
                        ws.write_merge(fil + 5, fil + 5, 0, 1, u'DIRECTOR(A) DE CARRERA', normalsinborde)
                        wb.save(response)
                        return response
                except Exception as es:
                    pass

            elif action == 'excelactatribunal':
                try:
                    __author__ = 'Unemi'
                    style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
                    style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
                    style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
                    title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                    style1 = easyxf(num_format_str='D-MMM-YY')
                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('exp_xls_post_part')
                    ws.write_merge(1, 1, 0, 5, u'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                    ws.write_merge(2, 2, 0, 5, u'VICERRECTORADO ACADÉMICO Y DE INVESTIGACIÓN', title)
                    ws.write_merge(3, 3, 0, 5, u'GESTIÓN TÉCNICA ACADÉMICA', title)
                    ws.write_merge(4, 4, 0, 5, u'PROCESO DE  TITULACIÓN', title)
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=Actas' + random.randint(1, 10000).__str__() + '.xls'
                    fech_ini = request.GET['fechainicio']
                    fech_fin = request.GET['fechafin']
                    columns = [
                        (u"N", 3000),
                        (u"CARRERA", 7000),
                        (u"TEMA", 10000),
                        (u"ALTERNATIVA TITULACIÓN", 7000),
                        (u"HORA INICIO", 4000),
                        (u"FECHA", 4000),
                        (u"N°", 4000),
                        (u"INTEGRANTES", 20000),
                        (u"TUTOR", 7000),
                        (u"PRESIDENTE", 7000),
                        (u"SECRETARIO", 7000),
                        (u"DELEGADO", 7000),
                        (u"CODIGO DE ACTA", 20000),
                        (u"ESTADO", 3000),
                        (u"NOTA FINAL", 20000),
                        (u"PERÍODO", 20000),
                        (u"ESTADO ARCHIVO FINAL", 10000),
                    ]

                    row_num = 5
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], font_style)
                        ws.col(col_num).width = columns[col_num][1]
                    date_format = xlwt.XFStyle()
                    date_format.num_format_str = 'yyyy-mm-dd'
                    date_formatreverse = xlwt.XFStyle()
                    date_formatreverse.num_format_str = 'dd/mm/yyyy'
                    # ComplexivoDetalleGrupo
                    # listadoactas = ComplexivoDetalleGrupo.objects.values_list('grupo').filter(fechaacta__gte=fech_ini, fechaacta__lte=fech_fin, actatribunalgenerada=True, status=True).distinct()
                    # grupos = ComplexivoGrupoTematica.objects.filter(pk__in=listadoactas, status=True)
                    grupos = ComplexivoGrupoTematica.objects.filter(fechadefensa__gte=fech_ini, fechadefensa__lte=fech_fin, status=True, activo=True)
                    row_num = 6
                    i = 0
                    for grupo in grupos:
                        campo1 = grupo.alternativa().carrera
                        campo2 = grupo.subtema
                        campo3 = grupo.alternativa()
                        campo4 = grupo.horadefensa
                        campo5 = grupo.fechadefensa
                        estadointegrante = ""
                        integrante = ""
                        codigoacta = ""
                        notafinal = ""
                        for inte in grupo.complexivodetallegrupo_set.filter(status=True):
                            apto = "NO APTO REQUISITOS"
                            if inte.matriculaaptahistorico:
                                apto = "APTO REQUISITOS"
                            estadointegrante = estadointegrante + u" • %s " % inte.matricula + ' - MATRÍCULA([' + str(inte.matricula.get_estado_display()) + '] ' + ' - [' + str(inte.matricula.get_cumplerequisitos_display()) + ']) ' + ' - GRUPO([' + str(inte.get_estadotribunal_display()) + '] ' + ' - [' + str(apto) + ']) ' + '\n'
                            integrante = integrante + u" • %s " % inte.matricula + ' - [' + str(inte.matricula.alternativa.paralelo) + '] ' + '\n'
                            dia = '0'
                            mes = '0'
                            anio = '0'
                            if inte.fechaacta:
                                dia = str(inte.fechaacta.day)
                                mes = str(inte.fechaacta.month)
                                anio = str(inte.fechaacta.year)
                                if inte.fechaacta.day < 10:
                                    dia = '0' + str(inte.fechaacta.day)
                                if inte.fechaacta.month < 10:
                                    mes = '0' + str(inte.fechaacta.month)
                            codigoacta = codigoacta + u" • %s - " % inte.matricula + ' - SUS-' + str(inte.numeroacta) + '-' + dia + mes + anio + '\n'
                            if inte.matricula.alternativa.tipotitulacion.tipo == 1:
                                notafinal = notafinal + u" • %s - " % inte.matricula + str(inte.matricula.notapropuesta(inte.grupo)) + '\n'
                            if inte.matricula.alternativa.tipotitulacion.tipo == 2:
                                notafinal = notafinal + u" • %s - " % inte.matricula + str(inte.matricula.notafinalcomplexivo(inte.grupo)) + '\n'

                        campo6 = integrante
                        camponro = grupo.complexivodetallegrupo_set.filter(status=True).count()
                        campo7 = ''
                        campo8 = ''
                        campo9 = ''
                        campo14 = ''
                        if grupo.tematica.tutor:
                            campo14 = grupo.tematica.tutor.__str__()
                        if grupo.presidentepropuesta:
                            campo7 = grupo.presidentepropuesta.persona
                        if grupo.secretariopropuesta:
                            campo8 = grupo.secretariopropuesta.persona
                        if grupo.delegadopropuesta:
                            campo9 = grupo.delegadopropuesta.persona
                        campo10 = codigoacta
                        campo11 = estadointegrante
                        campo12 = notafinal
                        campo13 = grupo.tematica.periodo.nombre
                        campo15 = ''
                        if grupo.subirarchivofinalgrupo:
                            if grupo.archivofinalgrupo:
                                campo15 = grupo.get_estadoarchivofinalgrupo_display()
                            else:
                                campo15 = 'PENDIENTE'

                        i += 1

                        ws.write(row_num, 0, i, font_style2)
                        ws.write(row_num, 1, u'%s' % campo1, font_style2)
                        ws.write(row_num, 2, u'%s' % campo2, font_style2)
                        ws.write(row_num, 3, u'%s' % campo3, font_style2)
                        ws.write(row_num, 4, u'%s' % campo4, font_style2)
                        ws.write(row_num, 5, campo5, date_format)
                        ws.write(row_num, 6, u'%s' % camponro, font_style2)
                        ws.write(row_num, 7, u'%s' % campo6, font_style2)
                        ws.write(row_num, 8, u'%s' % campo14, font_style2)
                        ws.write(row_num, 9, u'%s' % campo7, font_style2)
                        ws.write(row_num, 10, u'%s' % campo8, font_style2)
                        ws.write(row_num, 11, u'%s' % campo9, font_style2)
                        ws.write(row_num, 12, u'%s' % campo10, font_style2)
                        ws.write(row_num, 13, u'%s' % campo11, font_style2)
                        ws.write(row_num, 14, u'%s' % campo12, font_style2)
                        ws.write(row_num, 15, u'%s' % campo13, font_style2)
                        ws.write(row_num, 16, u'%s' % campo15, font_style2)

                        row_num += 1
                    wb.save(response)
                    return response
                except Exception as ex:
                    pass

            elif action == 'cambiartutor':
                try:
                    tematica = ComplexivoTematica.objects.get(pk=request.GET['id'])
                    data['tematica'] = tematica
                    lista = tematica.tematica.participantetematica_set.values_list("id", "participante__persona__nombres", "participante__persona__apellido1", "participante__persona__apellido2", flat=False).filter(status=True, participante__persona__isnull=False)
                    profesores = []
                    for lis in lista:
                        profesores.append([lis[0], "%s %s %s" % (lis[2], lis[3], lis[1])])
                    data['profesores'] = profesores
                    data['profesorselect'] = tematica.tutor.id
                    template = get_template("adm_complexivotematica/cambiartutor.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": "ok", 'data': json_content})
                except Exception as es:
                    pass

            elif action == 'listadofirmas':
                try:
                    data['title'] = u'Listado Firmas'
                    data['periodotitulacion'] = periodotitulacion = PeriodoGrupoTitulacion.objects.get(pk=int(request.GET['idperiodo']))
                    data['listadofirmas'] = periodotitulacion.firmaperiodogrupotitulacion_set.filter(status=True)
                    return render(request, "adm_complexivotematica/listadofirmas.html", data)
                except Exception as ex:
                    pass

            elif action == 'addfirma':
                try:
                    data['title'] = u'Adicionar firma'
                    data['periodotitulacion'] = PeriodoGrupoTitulacion.objects.get(pk=int(request.GET['idperiodotitulacion']))
                    form = FirmasPeriodoGrupoTitulacionForm()
                    data['form'] = form
                    return render(request, "adm_complexivotematica/addfirma.html", data)
                except Exception as ex:
                    pass

            elif action == 'reportetutoresgrupo_exc':
                try:
                    fechaactual = datetime.now().date()
                    periodogrupo = PeriodoGrupoTitulacion.objects.get(pk=request.GET['idper'])
                    __author__ = 'Unemi'
                    # style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
                    # style_nb = easyxf('font: name Times New Roman, color-index blue, bold on',num_format_str='#,##0.00')
                    # style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
                    # title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                    # style1 = easyxf(num_format_str='D-MMM-YY')
                    # font_style = XFStyle()
                    # font_style.font.bold = True
                    # font_style2 = XFStyle()
                    # font_style2.font.bold = False
                    output = io.BytesIO()
                    workbook = xlsxwriter.Workbook(output)
                    ws = workbook.add_worksheet('exp_xls_post_part')
                    # wb = Workbook(encoding='utf-8')
                    # ws = wb.add_sheet('exp_xls_post_part')
                    title = workbook.add_format({
                        'bold': 1,
                        'border': 1,
                        'align': 'center',
                        'valign': 'vcenter',
                        'bg_color': 'silver',
                        'text_wrap': 1,
                        'font_size': 10})

                    formatocelda = workbook.add_format({'border': 1})

                    formatotitulo = workbook.add_format(
                        {'align': 'center', 'valign': 'vcenter', 'bold': 1, 'font_size': 14})

                    ws.merge_range(0, 0, 0, 5, 'UNIVERSIDAD ESTATAL DE MILAGRO', formatotitulo)
                    ws.merge_range(1, 1, 0, 5, 'DIRECCIÓN DE GESTIÓN Y SERVICIOS ACADÉMICOS', formatotitulo)
                    ws.merge_range(2, 2, 0, 5, 'GESTIÓN TÉCNICA ACADÉMICA', formatotitulo)
                    ws.merge_range(3, 3, 0, 5, 'PROCESO DE TITULACIÓN', formatotitulo)
                    ws.merge_range(4, 4, 0, 5, periodogrupo.nombre, formatotitulo)
                    ws.merge_range(5, 5, 0, 5, 'REPORTE TUTORES CARGA HORARIA / GRUPOS ASIGNADOS', formatotitulo)
                    ws.merge_range(6, 6, 0, 5, 'FECHA: ' + str(fechaactual), formatotitulo)
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=exp_xls_post_part_' + random.randint(1, 10000).__str__() + '.xls'

                    columns = [
                        (u"FACULTAD", 15000),
                        (u"CARRERA", 15000),
                        (u"DOCENTE", 15000),
                        (u"DIRECCION Y TUTORIA DE TRABAJOS PARA LA OBTENCIÓN DEL TITULO DE GRADO CONFORMACION DE TRIBUNAL DE SUSTENTACIÓN EN MODALIDAD PRESENCIAL Y O EN LINEA", 3000),
                        (u"GRUPOS TITULACIÓN", 3000),
                        (u"OBSERVACIÓN", 15000),
                    ]
                    row_num = 8
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0])
                        ws.set_column(col_num, col_num, columns[col_num][1])
                    # for col_num in range(len(columns)):
                    #     ws.write(row_num, col_num, columns[col_num][0], formatotitulo)
                    #     ws.set_column(col_num, col_num, columns[col_num][1])
                    data = {}
                    listamatriz = []
                    listamatrizcoordinacion = []
                    data['fechaactual'] = datetime.now()
                    idper = int(request.GET['idper'])
                    codifacu = int(request.GET['codifacu'])
                    codicarr = int(request.GET['codicarr'])

                    listadofirmas = periodogrupo.firmaperiodogrupotitulacion_set.filter(status=True).order_by('tipofirma')
                    if idper > 0 and codifacu == 0 and codicarr == 0:
                        tematicas = ComplexivoTematica.objects.values_list('tutor__participante__persona__id', 'tutor__participante__persona__apellido1', 'tutor__participante__persona__apellido2', 'tutor__participante__persona__nombres', 'carrera__id', 'carrera__nombre', 'carrera__coordinacion__id', 'carrera__coordinacion__nombre').filter(periodo_id=idper, status=True, tutor__status=True).distinct().order_by('carrera__nombre', 'tutor__participante__persona__apellido1')
                    if codifacu > 0 and codicarr == 0:
                        tematicas = ComplexivoTematica.objects.values_list('tutor__participante__persona__id', 'tutor__participante__persona__apellido1', 'tutor__participante__persona__apellido2', 'tutor__participante__persona__nombres', 'carrera__id', 'carrera__nombre', 'carrera__coordinacion__id', 'carrera__coordinacion__nombre').filter(carrera__coordinacion__id=codifacu, periodo_id=idper, status=True, tutor__status=True).distinct().order_by('carrera__nombre', 'tutor__participante__persona__apellido1')
                    if codicarr > 0:
                        tematicas = ComplexivoTematica.objects.values_list('tutor__participante__persona__id', 'tutor__participante__persona__apellido1', 'tutor__participante__persona__apellido2', 'tutor__participante__persona__nombres', 'carrera__id', 'carrera__nombre', 'carrera__coordinacion__id', 'carrera__coordinacion__nombre').filter(carrera__id=codicarr, periodo_id=idper, status=True, tutor__status=True).distinct().order_by('carrera__nombre', 'tutor__participante__persona__apellido1')

                    totalhorasgrupo = 0
                    for tematica in tematicas:
                        horasactividad = 0
                        if DetalleDistributivo.objects.filter(distributivo__profesor__persona__id=tematica[0], distributivo__periodo=periodo, criteriodocenciaperiodo__criterio__id__in=[29, 99], status=True):
                            distri = DetalleDistributivo.objects.get(distributivo__profesor__persona__id=tematica[0], distributivo__periodo=periodo, criteriodocenciaperiodo__criterio__id__in=[29, 99], status=True)
                            horasactividad = distri.horas
                        totalhoras = ComplexivoGrupoTematica.objects.filter(status=True, tematica__carrera_id=tematica[4], tematica__periodo_id=idper, tematica__tutor__participante__persona_id=tematica[0])
                        totalhorasgrupo = totalhorasgrupo + totalhoras.count()
                        listamatriz.append([tematica[0], tematica[1], tematica[2], tematica[3], tematica[4], tematica[5],
                                            totalhoras.count(), tematica[6], tematica[7], horasactividad])
                    listacoordinaciones = Coordinacion.objects.filter(pk__in=tematicas.values_list('carrera__coordinacion__id', flat=True)).distinct()
                    listacarreras = None
                    if codicarr > 0:
                        listacarreras = Carrera.objects.filter(pk__in=tematicas.values_list('carrera__id', flat=True)).distinct()
                    listadotematicas = listamatriz
                    sede = Sede.objects.get(pk=1)
                    totalhorasgrupo = totalhorasgrupo
                    row_num = 9
                    for tematica in listadotematicas:
                        i = 0
                        campo1 = tematica[8]
                        campo2 = tematica[5]
                        campo3 = tematica[1] + ' ' + tematica[2] + ' ' + tematica[3]
                        campo4 = tematica[9]
                        campo5 = tematica[6]
                        campo6 = ''
                        ws.write(row_num, 0, campo1, formatocelda)
                        ws.write(row_num, 1, campo2, formatocelda)
                        ws.write(row_num, 2, campo3, formatocelda)
                        ws.write(row_num, 3, campo4, formatocelda)
                        ws.write(row_num, 4, campo5, formatocelda)
                        ws.write(row_num, 5, campo6, formatocelda)
                        row_num += 1
                    row_num += 3
                    ws.write(row_num, 0, 'ROL/CARGO', formatocelda)
                    ws.write(row_num, 1, 'FIRMA', formatocelda)
                    row_num += 1
                    for firma in listadofirmas:
                        ws.write(row_num, 0, firma.get_tipofirma_display(), formatocelda)
                        row_num += 1
                        ws.write(row_num, 0, firma.persona.apellido1 + ' ' + firma.persona.apellido2 + ' ' + firma.persona.nombres, formatocelda)
                        row_num += 1
                        ws.write(row_num, 0, firma.persona.mi_cargo_actualadm().denominacionpuesto.descripcion, formatocelda)
                        row_num += 2
                    for firmacoordinacion in listacoordinaciones:
                        decano = firmacoordinacion.responsable_periododos(periodo, 1)
                        ws.write(row_num, 0, 'Recibido por:', formatocelda)
                        row_num += 1
                        ws.write(row_num, 0, decano.persona.apellido1 + ' ' + decano.persona.apellido2 + ' ' + decano.persona.nombres, formatocelda)
                        row_num += 1
                        ws.write(row_num, 0, decano.persona.mi_cargo_actualadm().denominacionpuesto.descripcion, formatocelda)
                        row_num += 1
                        ws.write(row_num, 0, firmacoordinacion.nombre, formatocelda)
                        row_num += 2
                    if listacarreras:
                        for firmacarrera in listacarreras:
                            directorcarrera = firmacarrera.coordinador(periodo, 1)
                            ws.write(row_num, 0, 'Recibido por:', formatocelda)
                            row_num += 1
                            ws.write(row_num, 0, directorcarrera.persona.apellido1 + ' ' + directorcarrera.persona.apellido2 + ' ' + directorcarrera.persona.nombres, formatocelda)
                            row_num += 1
                            ws.write(row_num, 0, directorcarrera.persona.mi_cargo_actualadm().denominacionpuesto.descripcion, formatocelda)
                            row_num += 1
                            ws.write(row_num, 0, firmacarrera.nombre, formatocelda)
                            row_num += 2
                    workbook.close()
                    output.seek(0)
                    # Set up the Http response.
                    filename = 'horascumplidaspracvinc.xlsx'
                    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    response['Content-Disposition'] = 'attachment; filename=%s' % filename

                    return response
                except Exception as ex:
                    pass

            elif action == 'editartutor':
                try:
                    data['title'] = u"Editar tutor"
                    data['grupo'] = grupo = ComplexivoGrupoTematica.objects.get(id=request.GET['id'])
                    form = ParticipanteTematicaForm(initial={'tematica': grupo.tematica})
                    form.editar(grupo)
                    data['form'] = form
                    return render(request, "adm_grupoinvestigacion/editartutor.html", data)
                except Exception as ex:
                    pass

            elif action == 'reportetplanestudiantes':
                try:
                    fechaactual = datetime.now().date()
                    periodogrupo = PeriodoGrupoTitulacion.objects.get(pk=request.GET['idper'])
                    complexivotematica = ComplexivoTematica.objects.filter(status=True, periodo=periodogrupo, tutor__isnull=False)

                    #  creacion de archivo excel
                    __author__ = 'Unemi'
                    output = io.BytesIO()
                    workbook = xlsxwriter.Workbook(output)
                    ws = workbook.add_worksheet('PLAN {}'.format(periodogrupo.fechafin.year))
                    title = workbook.add_format({
                        'bold': 1,
                        'border': 1,
                        'align': 'center',
                        'valign': 'vcenter',
                        'bg_color': 'silver',
                        'text_wrap': 1,
                        'font_size': 10})

                    formatocelda = workbook.add_format({'border': 1})

                    formatotitulo = workbook.add_format(
                        {'align': 'center', 'valign': 'vcenter', 'bold': 1, 'font_size': 14})

                    formatocabecera = workbook.add_format(
                        {'align': 'center', 'valign': 'vcenter', 'bold': 1})

                    columns = [
                        (u"No.", 5),
                        (u"No. INTEGRANTES", 15),
                        (u"FACULTAD", 60),
                        (u"CARRERA", 60),
                        (u"TEMA/VARIABLE", 100),
                        (u"ALTERNATIVA TITULACIÓN", 30),
                        (u"HORA INICIO", 15),
                        (u"HORA FIN", 15),
                        (u"FECHA", 15),
                        (u"INTEGRANTES", 80),
                        (u"DOCENTE TUTOR", 40),
                        (u"PRESIDENTE", 15),
                        (u"SECRETARIO", 15),
                        (u"DELEGADO", 15),
                        (u"GTA ASIGNACION", 15),
                        (u"ESTADO FINAL PREVIO SUSTENTACION", 15),
                        (u"PERIODO", 15),
                        (u"OBSERVACIÓN", 15),
                        (u"CONVOCAR / NO CONVOCAR", 15)
                    ]
                    last_column = len(columns)
                    ws.merge_range(0, 0, 0, last_column, 'UNIVERSIDAD ESTATAL DE MILAGRO', formatotitulo)
                    ws.merge_range(1, 1, 1, last_column, 'DIRECCIÓN DE GESTIÓN Y SERVICIOS ACADÉMICOS', formatotitulo)
                    ws.merge_range(2, 1, 2, last_column, 'GESTIÓN TÉCNICA ACADÉMICA', formatotitulo)
                    ws.merge_range(3, 1, 3, last_column, 'PROCESO DE TITULACIÓN', formatotitulo)
                    ws.merge_range(4, 1, 4, last_column, periodogrupo.nombre, formatotitulo)
                    ws.merge_range(5, 1, 5, last_column, 'REPORTE PLANIFICACIÓN ESTUDIANTES', formatotitulo)
                    ws.merge_range(6, 1, 6, last_column, 'FECHA: ' + str(fechaactual), formatotitulo)
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=reporte_planificacion_estudiante_' + random.randint(1, 10000).__str__() + '.xls'

                    row_num = 7
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], formatocabecera)
                        ws.set_column(col_num, col_num, columns[col_num][1])
                    data = {}

                    row_num = 8
                    for tematica in complexivotematica:
                        for grupo in tematica.complexivogrupotematica_set.filter(status=True, activo=True).order_by('id'):
                            part = ''
                            part = ''
                            for p in grupo.participantes():
                                part = part + u"• %s" % p.matricula + '-[' + str(
                                    p.matricula.alternativa.paralelo) + '] ' + '\n'
                                # part += str(p.matricula.inscripcion.persona) + ','
                            campo1 = grupo.pk
                            campo2 = len(grupo.participantes())
                            campo3 = tematica.carrera.mi_coordinacion()
                            campo4 = tematica.carrera.nombre
                            campo5 = tematica.tematica.tema
                            campo6 = str(grupo.participantes()[0].matricula.alternativa)
                            campo7 = ''
                            campo8 = ''
                            campo9 = ''
                            campo10 = part
                            campo11 = str(tematica.tutor)
                            campo12 = ''
                            campo13 = ''
                            campo14 = ''
                            campo15 = ''
                            campo16 = grupo.estado_propuesta().get_estado_display() if grupo.estado_propuesta() else ""
                            campo17 = str(tematica.periodo.nombre)
                            campo18 = ''
                            campo19 = ''
                            ws.write(row_num, 0, campo1, formatocelda)
                            ws.write(row_num, 1, campo2, formatocelda)
                            ws.write(row_num, 2, campo3, formatocelda)
                            ws.write(row_num, 3, campo4, formatocelda)
                            ws.write(row_num, 4, campo5, formatocelda)
                            ws.write(row_num, 5, campo6, formatocelda)
                            ws.write(row_num, 6, campo7, formatocelda)
                            ws.write(row_num, 7, campo8, formatocelda)
                            ws.write(row_num, 8, campo9, formatocelda)
                            ws.write(row_num, 9, campo10, formatocelda)
                            ws.write(row_num, 10, campo11, formatocelda)
                            ws.write(row_num, 11, campo12, formatocelda)
                            ws.write(row_num, 12, campo13, formatocelda)
                            ws.write(row_num, 13, campo14, formatocelda)
                            ws.write(row_num, 14, campo15, formatocelda)
                            ws.write(row_num, 15, campo16, formatocelda)
                            ws.write(row_num, 16, campo17, formatocelda)
                            ws.write(row_num, 17, campo18, formatocelda)
                            ws.write(row_num, 18, campo19, formatocelda)
                            row_num += 1

                    workbook.close()
                    output.seek(0)
                    filename = 'reporte_planificacion_estudiante_' + random.randint(1, 10000).__str__() + '.xlsx'
                    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    response['Content-Disposition'] = 'attachment; filename=%s' % filename
                    return response
                except Exception as ex:
                    print(ex)

            elif action == 'reporteextension':
                try:
                    fechaactual = datetime.now().date()
                    periodogrupo = PeriodoGrupoTitulacion.objects.get(pk=request.GET['idper'])
                    grupos = ComplexivoGrupoTematica.objects.filter(status=True, tematica__periodo=periodogrupo, subirpropuesta=True, tematica__tutor__isnull=False)

                    #  creacion de archivo excel
                    __author__ = 'Unemi'
                    output = io.BytesIO()
                    workbook = xlsxwriter.Workbook(output)
                    ws = workbook.add_worksheet('PLAN {}'.format(periodogrupo.fechafin.year))
                    title = workbook.add_format({
                        'bold': 1,
                        'border': 1,
                        'align': 'center',
                        'valign': 'vcenter',
                        'bg_color': 'silver',
                        'text_wrap': 1,
                        'font_size': 10})

                    formatocelda = workbook.add_format({'border': 1})

                    formatotitulo = workbook.add_format(
                        {'align': 'center', 'valign': 'vcenter', 'bold': 1, 'font_size': 14})

                    formatocabecera = workbook.add_format(
                        {'align': 'center', 'valign': 'vcenter', 'bold': 1})

                    columns = [
                        (u"Cod. grupo", 10),
                        (u"Estudiantes", 100),
                        (u"Tutor", 60),
                        (u"Periodo", 60),
                        (u"Fecha inicio estudiante", 60),
                        (u"Fecha fin estudiante", 60),
                        (u"Fecha inicio tutor", 60),
                        (u"Fecha fin tutor", 60),
                    ]
                    last_column = len(columns)
                    ws.merge_range(0, 0, 0, last_column, 'UNIVERSIDAD ESTATAL DE MILAGRO', formatotitulo)
                    ws.merge_range(1, 1, 1, last_column, 'DIRECCIÓN DE GESTIÓN Y SERVICIOS ACADÉMICOS', formatotitulo)
                    ws.merge_range(2, 1, 2, last_column, 'GESTIÓN TÉCNICA ACADÉMICA', formatotitulo)
                    ws.merge_range(3, 1, 3, last_column, 'PROCESO DE TITULACIÓN', formatotitulo)
                    ws.merge_range(4, 1, 4, last_column, periodogrupo.nombre, formatotitulo)
                    ws.merge_range(5, 1, 5, last_column, 'REPORTE PLANIFICACIÓN ESTUDIANTES', formatotitulo)
                    ws.merge_range(6, 1, 6, last_column, 'FECHA: ' + str(fechaactual), formatotitulo)
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=reporte_extension_' + random.randint(1, 10000).__str__() + '.xls'

                    row_num = 7
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], formatocabecera)
                        ws.set_column(col_num, col_num, columns[col_num][1])
                    data = {}

                    row_num = 8
                    for grupo in grupos:
                        part = ''
                        for p in grupo.participantes():
                            part += str(p.matricula.inscripcion.persona) + ','

                        ws.write(row_num, 0, str(grupo.pk), formatocelda)
                        ws.write(row_num, 1, part, formatocelda)
                        ws.write(row_num, 2, str(grupo.tematica.tutor), formatocelda)
                        ws.write(row_num, 3, str(grupo.tematica.periodo.nombre), formatocelda)
                        ws.write(row_num, 4, str(grupo.fechasubirpropuesta), formatocelda)
                        ws.write(row_num, 5, str(grupo.fecha_extension_alu()), formatocelda)
                        ws.write(row_num, 6, str(grupo.fechasubirpropuesta), formatocelda)
                        ws.write(row_num, 7, str(grupo.fecha_extension_doc()), formatocelda)
                        row_num += 1

                    workbook.close()
                    output.seek(0)
                    filename = 'reporte_extension_' + random.randint(1, 10000).__str__() + '.xlsx'
                    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    response['Content-Disposition'] = 'attachment; filename=%s' % filename
                    return response
                except Exception as ex:
                    print(ex)

            elif action == 'addtiempo':
                try:
                    form = ConfiguracionComplexivoHabilitaPropuestaForm()
                    data['form2'] = form
                    template = get_template("adm_complexivotematica/modal/formtiempop.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'edittiempo':
                try:
                    data['id'] = request.GET['id']
                    data['fitro'] = filtro = ConfiguracionComplexivoHabilitaPropuesta.objects.get(pk=request.GET['id'])
                    data['form2'] = ConfiguracionComplexivoHabilitaPropuestaForm(initial=model_to_dict(filtro))
                    template = get_template("adm_complexivotematica/modal/formtiempop.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'requisitosmateriatitulacion':
                try:
                    data['title'] = u'Requisitos titulacion'
                    data['materia'] = materia = Materia.objects.get(pk=int(encrypt(request.GET['id'])))
                    requisitos = materia.requisitomateriaunidadintegracioncurricular_set.filter(status=True).order_by('requisito__nombre')
                    existegraduado = MateriaTitulacion.objects.values("id").filter(materiaasignada__materia=materia, estadograduado=True).exists()
                    if (not requisitos and not existegraduado) or 'sync' in request.GET:
                        response = llenar_requisitostitulacion(materia, request)
                        if response['resp'] == 'error':
                            return HttpResponseRedirect("/adm_complexivotematica?info={}".format(response['msg']))
                    data['existegraduado'] = existegraduado
                    data['atras'] = '/adm_complexivotematica?action=asignaturastitulacion'
                    data['listadorequisitos'] = requisitos
                    # return render(request, "adm_complexivotematica/requisitosmateriatitulacion.html", data)
                    return render(request, "niveles/requisitosmateriatitulacion.html", data)

                except Exception as ex:
                    pass

            elif action == 'listadodocentefirma':
                try:
                    data['title'] = u'Listado de docentes a firmar acta titulación'
                    data['grupo'] = grupo = GrupoTitulacionIC.objects.get(pk=int(encrypt(request.GET['id'])))
                    existegraduado = False
                    if MateriaTitulacion.objects.values("id").filter(materiaasignada__materia=grupo.materia, estadograduado=True).exists():
                        existegraduado = True
                    data['existegraduado'] = existegraduado
                    data['listadogrupofirmas'] = grupo.grupofirma_set.filter(status=True).order_by('id')
                    return render(request, "adm_complexivotematica/listadodocentefirma.html", data)
                except Exception as ex:
                    pass

            elif action == 'adddocentefirma':
                try:
                    data['title'] = u'Adicionar Docente'
                    data['idgrupo'] = request.GET['id']
                    form = FirmaGrupoTitulacionForm()
                    data['form'] = form
                    template = get_template('adm_complexivotematica/adddocentefirma.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'buscardocente':
                try:
                    if 'q' in request.GET:
                        search = request.GET['q'].upper().strip()
                        ss = search.split(' ')
                        if len(ss) == 1:
                            query = Profesor.objects.filter(Q(persona__nombres__icontains=search) |
                                                            Q(persona__apellido1__icontains=search) |
                                                            Q(persona__apellido2__icontains=search) |
                                                            Q(persona__cedula__icontains=search) |
                                                            Q(persona__pasaporte__icontains=search)).distinct()
                        elif len(ss) == 2:
                            query = Profesor.objects.filter(Q(persona__apellido1__icontains=ss[0]) & Q(persona__apellido2__icontains=ss[1])).distinct()
                        elif len(ss) == 3:
                            query = Profesor.objects.filter(Q(persona__apellido1__icontains=ss[0]) | Q(persona__apellido2__icontains=ss[1]) | Q(persona__apellido2__icontains=ss[2])).distinct()
                        else:
                            query = Profesor.objects.filter(Q(persona__apellido1__icontains=ss[0]) | Q(persona__apellido2__icontains=ss[1]) | Q(persona__apellido2__icontains=ss[2]) | Q(persona__apellido2__icontains=ss[3])).distinct()
                    else:
                        query = Profesor.objects.filter(status=True).distinct()
                    data = {"results": [{"id": x.id, "name": x.persona.nombre_completo_inverso()} for x in query]}
                    return JsonResponse(data)
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": 'Error al obtener los datos.'})

            elif action == 'buscarpersonas':
                try:
                    q = request.GET['q'].upper().strip()
                    s = q.split(" ")
                    querybase = Persona.objects.filter(status=True, ).order_by('apellido1')
                    if len(s) == 1:
                        per = querybase.filter((Q(nombres__icontains=q) | Q(apellido1__icontains=q) | Q(cedula__icontains=q) | Q(apellido2__icontains=q) | Q(cedula__contains=q)),
                                               Q(status=True)).distinct()[:15]
                    elif len(s) == 2:
                        per = querybase.filter((Q(apellido1__contains=s[0]) & Q(apellido2__contains=s[1])) |
                                               (Q(nombres__icontains=s[0]) & Q(nombres__icontains=s[1])) |
                                               (Q(nombres__icontains=s[0]) & Q(apellido1__contains=s[1]))).filter(status=True).distinct()[:15]
                    else:
                        per = querybase.filter((Q(nombres__contains=s[0]) & Q(apellido1__contains=s[1]) & Q(apellido2__contains=s[2])) |
                                               (Q(nombres__contains=s[0]) & Q(nombres__contains=s[1]) & Q(apellido1__contains=s[2]))).filter(status=True).distinct()[:15]
                    data = {"result": "ok", "results": [{"id": x.id, "name": "{} - {}".format(x.cedula, x.nombre_completo())} for x in per]}
                    return JsonResponse(data)
                except Exception as ex:
                    pass

            elif action == 'configuracalificacion':
                try:
                    data['title'] = u'Rúbrica de titulación'
                    data['materia'] = materia = Materia.objects.get(pk=int(encrypt(request.GET['id'])))
                    tienegraduados = False
                    if MateriaTitulacion.objects.values("id").filter(materiaasignada__materia=materia, estadograduado=True).exists():
                        tienegraduados = True
                    data['tienegraduados'] = tienegraduados
                    if materia.asignaturamalla.malla.carrera.id != 208:
                        if not materia.grupotitulacionic_set.filter(status=True).exists():
                            grupo = GrupoTitulacionIC(materia=materia)
                            grupo.save(request)
                        else:
                            grupo = GrupoTitulacionIC.objects.get(materia=materia)
                        listaprodecesoratitulacion = materia.asignaturamalla.asignaturamallapredecesora_set.filter(predecesora__validarequisitograduacion=True, status=True)
                        numorden = 0
                        for lpro in listaprodecesoratitulacion:
                            numorden += 1
                            if not MateriaGrupoTitulacion.objects.values('id').filter(grupo=grupo, asignaturamalla=lpro.predecesora, status=True).exists():
                                materiagrupo = MateriaGrupoTitulacion(grupo=grupo,
                                                                      asignaturamalla=lpro.predecesora,
                                                                      nombre='',
                                                                      puntaje=0,
                                                                      orden=numorden)
                                materiagrupo.save(request)
                        numorden += 1
                        if not MateriaGrupoTitulacion.objects.values('id').filter(grupo=grupo, asignaturamalla=materia.asignaturamalla, status=True).exists():
                            materiagrupo = MateriaGrupoTitulacion(grupo=grupo,
                                                                  asignaturamalla=materia.asignaturamalla,
                                                                  nombre='',
                                                                  puntaje=0,
                                                                  orden=numorden)
                            materiagrupo.save(request)
                    data['listado'] = MateriaGrupoTitulacion.objects.filter(grupo__materia_id=materia.id, status=True).order_by('orden')
                    return render(request, "adm_complexivotematica/configuracalificacion.html", data)
                except Exception as ex:
                    pass

            elif action == 'reportasignatura':
                try:
                    if 'idmalla' in request.GET:
                        conrequisitos=request.GET['conrequisitos']
                        periodo = request.session['periodo']
                        idmalla = int(encrypt(request.GET['idmalla']))
                        if int(conrequisitos) == 1:
                            if idmalla == 383:
                                asignaturas = AsignaturaMalla.objects.filter(malla_id=idmalla, status=True).order_by('malla__carrera__nombre')
                            else:
                                asignaturas = AsignaturaMalla.objects.filter(malla_id=idmalla,validarequisitograduacion=True,status=True).order_by('malla__carrera__nombre')
                            output = io.BytesIO()
                            workbook = xlsxwriter.Workbook(output)
                            formatorojo = workbook.add_format({'bold': True, 'font_color': 'red'})
                            for asignatura in asignaturas:
                                # requisitos = RequisitoIngresoUnidadIntegracionCurricular.objects.filter(status=True,asignaturamalla=asignatura)
                                materias = Materia.objects.filter(nivel__periodo_id=periodo, asignaturamalla=asignatura, status=True).order_by('paralelo')
                                if materias:
                                    for materia in materias:
                                        worksheet = workbook.add_worksheet()
                                        asignados = MateriaAsignada.objects.filter(status=True, materia=materia, retiramateria=False).order_by('matricula__inscripcion__persona__apellido1', 'matricula__inscripcion__persona__apellido2')
                                        worksheet.write(0, 0, 'Inscrito')
                                        worksheet.write(0, 1, 'Identificación')
                                        worksheet.write(0, 2, 'Correo Inst')
                                        worksheet.write(0, 3, 'Correo Personal')
                                        worksheet.write(0, 4, 'Materia')
                                        worksheet.write(0, 5, 'Paralelo')
                                        worksheet.write(0, 6, 'Facultad')
                                        worksheet.write(0, 7, 'Carrera materia')
                                        worksheet.write(0, 8, 'Carrera alumno')
                                        worksheet.write(0, 9, 'Malla')
                                        col = 10
                                        fil = 1
                                        requisitos = materia.requisitomateriaunidadintegracioncurricular_set.filter(activo=True, titulacion=True, status=True)
                                        for erequisito in requisitos:
                                            worksheet.write(0, col, str(erequisito.requisito))
                                            col += 1
                                        for asignado in asignados:
                                            inscripcion = asignado.matricula.inscripcion
                                            worksheet.write(fil, 0, str(inscripcion.persona))
                                            worksheet.write(fil, 1, str(inscripcion.persona.identificacion()))
                                            worksheet.write(fil, 2, str(inscripcion.persona.emailinst))
                                            worksheet.write(fil, 3, str(inscripcion.persona.email))
                                            worksheet.write(fil, 4, str(materia))
                                            worksheet.write(fil, 5, str(materia.paralelo))
                                            worksheet.write(fil, 6, str(materia.asignaturamalla.malla.carrera.mi_coordinacion()))
                                            worksheet.write(fil, 7, str(materia.asignaturamalla.malla.carrera))
                                            worksheet.write(fil, 8, str(inscripcion.carrera))
                                            worksheet.write(fil, 9, str(materia.asignaturamalla.malla))
                                            col = 10
                                            for erequisito in requisitos:
                                                cumple = erequisito.run(inscripcion.pk)
                                                estadocumple = 'NO CUMPLE'
                                                if cumple:
                                                    estadocumple = 'SI CUMPLE'
                                                    worksheet.write(fil, col, estadocumple)
                                                else:
                                                    worksheet.write(fil, col, estadocumple, formatorojo)
                                                col += 1
                                            fil += 1
                            workbook.close()
                            output.seek(0)
                            filename = 'reporte_requisitos' + random.randint(1, 10000).__str__() + '.xlsx'
                            response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                            response['Content-Disposition'] = 'attachment; filename=%s' % filename
                            return response
                        else:
                            output = io.BytesIO()
                            workbook = xlsxwriter.Workbook(output)
                            worksheet = workbook.add_worksheet()
                            formatorojo = workbook.add_format({'bold': True, 'font_color': 'red'})
                            if idmalla == 383:
                                asignados = MateriaAsignada.objects.filter(materia__nivel__periodo=periodo, status=True, materia__asignaturamalla__malla__id=idmalla, retiramateria=False).order_by('matricula__inscripcion__persona__apellido1', 'matricula__inscripcion__persona__apellido2')
                            else:
                                asignados = MateriaAsignada.objects.filter(materia__asignaturamalla__validarequisitograduacion=True, materia__nivel__periodo=periodo, status=True, materia__asignaturamalla__malla__id=idmalla, retiramateria=False).order_by('matricula__inscripcion__persona__apellido1', 'matricula__inscripcion__persona__apellido2')
                            worksheet.write(0, 0, 'Inscrito')
                            worksheet.write(0, 1, 'Identificación')
                            worksheet.write(0, 2, 'Correo Inst')
                            worksheet.write(0, 3, 'Correo Personal')
                            worksheet.write(0, 4, 'Materia')
                            worksheet.write(0, 5, 'Paralelo')
                            worksheet.write(0, 6, 'Facultad')
                            worksheet.write(0, 7, 'Carrera materia')
                            worksheet.write(0, 8, 'Carrera alumno')
                            worksheet.write(0, 9, 'Malla')
                            worksheet.write(0, 10, 'ESTADO')
                            worksheet.write(0, 11, 'APTO REQUISITOS')
                            fil = 1
                            for asignado in asignados:
                                cumple = asignado.matricula.inscripcion.valida_requisitos_complexivo(asignado.materia.id, False)

                                worksheet.write(fil, 0, str(asignado.matricula.inscripcion.persona))
                                worksheet.write(fil, 1, str(asignado.matricula.inscripcion.persona.identificacion()))
                                worksheet.write(fil, 2, str(asignado.matricula.inscripcion.persona.emailinst))
                                worksheet.write(fil, 3, str(asignado.matricula.inscripcion.persona.email))
                                worksheet.write(fil, 4, str(asignado.materia))
                                worksheet.write(fil, 5, str(asignado.materia.paralelo))
                                worksheet.write(fil, 6, str(asignado.materia.asignaturamalla.malla.carrera.mi_coordinacion()))
                                worksheet.write(fil, 7, str(asignado.materia.asignaturamalla.malla.carrera))
                                worksheet.write(fil, 8, str(asignado.matricula.inscripcion.carrera))
                                worksheet.write(fil, 9, str(asignado.materia.asignaturamalla.malla))
                                worksheet.write(fil, 10, str(asignado.estado.nombre))
                                estadoapto = 'NO CUMPLE REQUISITOS'
                                if cumple:
                                    estadoapto = 'SI CUMPLE REQUISITOS'
                                    worksheet.write(fil, 11, str(estadoapto))
                                else:
                                    worksheet.write(fil, 11, str(estadoapto), formatorojo)
                                fil += 1
                            workbook.close()
                            output.seek(0)
                            filename = 'reporte_requisitos' + random.randint(1, 10000).__str__() + '.xlsx'
                            response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                            response['Content-Disposition'] = 'attachment; filename=%s' % filename
                            return response
                    else:
                        listado1 = Materia.objects.values_list('asignaturamalla__malla_id', flat=True).filter(asignaturamalla__validarequisitograduacion=True, nivel__periodo=periodo, asignaturamalla__status=True, status=True).distinct()
                        listado2 = Materia.objects.values_list('asignaturamalla__malla_id', flat=True).filter(asignaturamalla__malla_id=383, nivel__periodo=periodo, asignaturamalla__status=True, status=True).distinct()
                        listadomalla = listado1 | listado2
                        data['mallas'] = Malla.objects.filter(pk__in=listadomalla, status=True).order_by('carrera__nombre')
                        template = get_template("adm_complexivotematica/modal/formreporte.html")
                        return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'reportasignaturanivel':
                try:
                    if 'idmalla' in request.GET:
                        idnivel=request.GET['idnivel']
                        periodo = request.session['periodo']
                        listarequi = request.GET['lista'].split(',')
                        idmalla = int(encrypt(request.GET['idmalla']))
                        output = io.BytesIO()
                        workbook = xlsxwriter.Workbook(output)
                        formatorojo = workbook.add_format({'bold': True, 'font_color': 'red'})
                        listadoalumnos = MateriaAsignada.objects.filter(status=True, materia__asignaturamalla__malla_id=idmalla, materia__nivel__periodo=periodo, materia__asignaturamalla__nivelmalla_id=idnivel, retiramateria=False).distinct('matricula__inscripcion__persona__apellido1', 'matricula__inscripcion__persona__apellido2', 'matricula__inscripcion__persona__nombres').order_by('matricula__inscripcion__persona__apellido1', 'matricula__inscripcion__persona__apellido2', 'matricula__inscripcion__persona__nombres')
                        worksheet = workbook.add_worksheet("Listado")

                        worksheet.write(0, 0, 'Inscrito')
                        worksheet.write(0, 1, 'Identificación')
                        worksheet.write(0, 2, 'Correo Inst')
                        worksheet.write(0, 3, 'Correo Personal')
                        worksheet.write(0, 4, 'Nivel')
                        worksheet.write(0, 5, 'Carrera alumno')
                        col = 6
                        requisitos = RequisitoTitulacionMalla.objects.filter(pk__in=listarequi, malla_id=idmalla, status=True).order_by('requisito__nombre')
                        for erequisito in requisitos:
                            worksheet.write(0, col, str(erequisito.requisito))
                            col += 1
                        fil = 2
                        for asignado in listadoalumnos:
                            inscripcion = asignado.matricula.inscripcion
                            worksheet.write(fil, 0, str(inscripcion.persona))
                            worksheet.write(fil, 1, str(inscripcion.persona.identificacion()))
                            worksheet.write(fil, 2, str(inscripcion.persona.emailinst))
                            worksheet.write(fil, 3, str(inscripcion.persona.email))
                            worksheet.write(fil, 4, str(asignado.materia.asignaturamalla.nivelmalla.nombre))
                            worksheet.write(fil, 5, str(inscripcion.carrera))
                            col = 6
                            for erequisito in requisitos:
                                cumple = erequisito.run(inscripcion.pk)
                                estadocumple = 'NO CUMPLE'
                                if cumple:
                                    estadocumple = 'SI CUMPLE'
                                    worksheet.write(fil, col, estadocumple)
                                else:
                                    worksheet.write(fil, col, estadocumple, formatorojo)
                                col += 1
                            fil += 1
                        workbook.close()
                        output.seek(0)
                        filename = 'reporte_requisitos' + random.randint(1, 10000).__str__() + '.xlsx'
                        response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                        response['Content-Disposition'] = 'attachment; filename=%s' % filename
                        return response
                    else:
                        listado = Materia.objects.values_list('asignaturamalla__malla_id', flat=True).filter(nivel__periodo=periodo, asignaturamalla__status=True, status=True).exclude(asignaturamalla__malla__carrera__coordinacion=9).distinct()
                        data['mallas'] = Malla.objects.filter(pk__in=listado, status=True).order_by('carrera__nombre')
                        template = get_template("adm_complexivotematica/modal/formreportexnivel.html")
                        return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'matriculadosintegracion':
                try:
                    periodo = request.session['periodo']
                    output = io.BytesIO()
                    workbook = xlsxwriter.Workbook(output)
                    worksheet = workbook.add_worksheet()
                    formatorojo = workbook.add_format({'bold': True, 'font_color': 'red'})
                    idmalla = 383
                    asignados1 = MateriaAsignada.objects.filter(materia__nivel__periodo=periodo, status=True, materia__asignaturamalla__malla__id=idmalla, retiramateria=False).order_by('matricula__inscripcion__persona__apellido1', 'matricula__inscripcion__persona__apellido2')
                    asignados2 = MateriaAsignada.objects.filter(materia__asignaturamalla__validarequisitograduacion=True, materia__nivel__periodo=periodo, status=True, retiramateria=False).exclude(materia__asignaturamalla__malla__id=idmalla).order_by('matricula__inscripcion__persona__apellido1', 'matricula__inscripcion__persona__apellido2')
                    asignados = asignados1 | asignados2

                    worksheet.write(0, 0, 'Inscrito')
                    worksheet.write(0, 1, 'Identificación')
                    worksheet.write(0, 2, 'Correo Inst')
                    worksheet.write(0, 3, 'Correo Personal')
                    worksheet.write(0, 4, 'Materia')
                    worksheet.write(0, 5, 'Paralelo')
                    worksheet.write(0, 6, 'Facultad')
                    worksheet.write(0, 7, 'Carrera materia')
                    worksheet.write(0, 8, 'Carrera alumno')
                    worksheet.write(0, 9, 'Malla')
                    worksheet.write(0, 10, 'ESTADO')
                    worksheet.write(0, 11, 'APTO REQUISITOS')
                    worksheet.write(0, 12, 'VECES MATRICULADO')
                    worksheet.write(0, 13, 'PERIODO ACADÉMICO')
                    worksheet.write(0, 14, 'NUMERO MEMORANDO')
                    worksheet.write(0, 15, 'NÚMERO DE ACTA')
                    fil = 1
                    for asignado in asignados:
                        cumple = asignado.matricula.inscripcion.valida_requisitos_complexivo(asignado.materia.id, False)
                        vecesmatriculado = asignado.matricula.inscripcion.historicorecordacademico_set.values("id"). \
                                               filter(asignatura=asignado.materia.asignaturamalla.asignatura). \
                                               exclude(noaplica=True).exclude(convalidacion=True). \
                                               exclude(homologada=True).distinct().count() + 1
                        numeromemo = 0
                        numacta = 0
                        if asignado.materiatitulacion_set.filter(status=True):
                            asignadotitulacion = asignado.materiatitulacion_set.filter(status=True)[0]
                            numeromemo = asignadotitulacion.numeromemo
                            numacta = 'SUS-%s-%s-%s' % (asignadotitulacion.materiaasignada.matricula.inscripcion.carrera.abrsustentacion, asignadotitulacion.numeroacta, asignadotitulacion.materiaasignada.materia.nivel.periodo.fin)

                        worksheet.write(fil, 0, str(asignado.matricula.inscripcion.persona))
                        worksheet.write(fil, 1, str(asignado.matricula.inscripcion.persona.identificacion()))
                        worksheet.write(fil, 2, str(asignado.matricula.inscripcion.persona.emailinst))
                        worksheet.write(fil, 3, str(asignado.matricula.inscripcion.persona.email))
                        worksheet.write(fil, 4, str(asignado.materia))
                        worksheet.write(fil, 5, str(asignado.materia.paralelo))
                        worksheet.write(fil, 6, str(asignado.materia.asignaturamalla.malla.carrera.mi_coordinacion()))
                        worksheet.write(fil, 7, str(asignado.materia.asignaturamalla.malla.carrera))
                        worksheet.write(fil, 8, str(asignado.matricula.inscripcion.carrera))
                        worksheet.write(fil, 9, str(asignado.materia.asignaturamalla.malla))
                        worksheet.write(fil, 10, str(asignado.estado.nombre))
                        estadoapto = 'NO CUMPLE REQUISITOS'
                        if cumple:
                            estadoapto = 'SI CUMPLE REQUISITOS'
                            worksheet.write(fil, 11, str(estadoapto))
                        else:
                            worksheet.write(fil, 11, str(estadoapto), formatorojo)
                        worksheet.write(fil, 12, vecesmatriculado)
                        worksheet.write(fil, 13, str(asignado.materia.nivel.periodo))
                        worksheet.write(fil, 14, str(numeromemo))
                        worksheet.write(fil, 15, str(numacta))
                        fil += 1
                    workbook.close()
                    output.seek(0)
                    filename = 'reporte_requisitos' + random.randint(1, 10000).__str__() + '.xlsx'
                    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    response['Content-Disposition'] = 'attachment; filename=%s' % filename
                    return response
                except Exception as ex:
                    pass

            elif action == 'sinactaestudiantes':
                try:
                    periodo = request.session['periodo']
                    output = io.BytesIO()
                    workbook = xlsxwriter.Workbook(output)
                    worksheet = workbook.add_worksheet()
                    formatorojo = workbook.add_format({'bold': True, 'font_color': 'red'})
                    idmalla = 383

                    if 'periodoid' in request.GET:
                        asignados1 = MateriaAsignada.objects.filter(materia__nivel__periodo=periodo, status=True, materia__asignaturamalla__malla__id=idmalla, retiramateria=False).order_by('matricula__inscripcion__persona__apellido1', 'matricula__inscripcion__persona__apellido2')
                        asignados2 = MateriaAsignada.objects.filter(materia__nivel__periodo=periodo, materia__asignaturamalla__validarequisitograduacion=True, status=True, retiramateria=False).exclude(materia__asignaturamalla__malla__id=idmalla).order_by('matricula__inscripcion__persona__apellido1', 'matricula__inscripcion__persona__apellido2')
                    else:
                        asignados1 = MateriaAsignada.objects.filter(status=True, materia__asignaturamalla__malla__id=idmalla, retiramateria=False).order_by('matricula__inscripcion__persona__apellido1', 'matricula__inscripcion__persona__apellido2')
                        asignados2 = MateriaAsignada.objects.filter(materia__asignaturamalla__validarequisitograduacion=True, status=True, retiramateria=False).exclude(materia__asignaturamalla__malla__id=idmalla).order_by('matricula__inscripcion__persona__apellido1', 'matricula__inscripcion__persona__apellido2')
                    asignados = asignados1 | asignados2
                    qsmatriculados = MateriaTitulacion.objects.filter(actacerrada=False, status=True, materiaasignada__in=list(asignados.values_list('id',flat=True))).order_by('materiaasignada__materia')

                    worksheet.write(0, 0, 'Periodo')
                    worksheet.write(0, 1, 'Facultad')
                    worksheet.write(0, 2, 'Carrera')
                    worksheet.write(0, 3, 'Materia')
                    worksheet.write(0, 4, 'Paralelo')
                    worksheet.write(0, 5, 'Inscrito')
                    worksheet.write(0, 6, 'Identificación')
                    worksheet.write(0, 7, 'Correo Inst')
                    worksheet.write(0, 8, 'Correo Personal')
                    worksheet.write(0, 9, 'Nota Final')
                    worksheet.write(0, 10, 'Asist')
                    worksheet.write(0, 11, 'Estado')
                    worksheet.write(0, 12, 'CUMPLE REQUISITOS')
                    fil = 1
                    for asignadomateria in qsmatriculados:
                        cumple = asignadomateria.materiaasignada.matricula.inscripcion.valida_requisitos_complexivo(asignadomateria.materiaasignada.materia.id, False)

                        worksheet.write(fil, 0, str(asignadomateria.materiaasignada.matricula.nivel.periodo))
                        worksheet.write(fil, 1, str(asignadomateria.materiaasignada.matricula.inscripcion.carrera.nombre))
                        worksheet.write(fil, 2, str(asignadomateria.materiaasignada.matricula.inscripcion.coordinacion.nombre))
                        worksheet.write(fil, 3, str(asignadomateria.materiaasignada.materia))
                        worksheet.write(fil, 4, str(asignadomateria.materiaasignada.materia.paralelo))
                        worksheet.write(fil, 5, str(asignadomateria.materiaasignada.matricula.inscripcion.persona))
                        worksheet.write(fil, 6, str(asignadomateria.materiaasignada.matricula.inscripcion.persona.identificacion()))
                        worksheet.write(fil, 7, str(asignadomateria.materiaasignada.matricula.inscripcion.persona.emailinst))
                        worksheet.write(fil, 8, str(asignadomateria.materiaasignada.matricula.inscripcion.persona.email))
                        worksheet.write(fil, 9, str(asignadomateria.materiaasignada.notafinal))
                        worksheet.write(fil, 10, str(asignadomateria.materiaasignada.asistenciafinal))
                        worksheet.write(fil, 11, str(asignadomateria.materiaasignada.estado))
                        estadoapto = 'NO CUMPLE REQUISITOS'
                        if cumple:
                            estadoapto = 'SI CUMPLE REQUISITOS'
                            worksheet.write(fil, 12, str(estadoapto))
                        else:
                            worksheet.write(fil, 12, str(estadoapto), formatorojo)
                        fil += 1
                    workbook.close()
                    output.seek(0)
                    filename = 'reporte_sin_actas_estudiantes' + random.randint(1, 10000).__str__() + '.xlsx'
                    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    response['Content-Disposition'] = 'attachment; filename=%s' % filename
                    return response
                except Exception as ex:
                    pass

            elif action == 'alumdisertacion':
                try:
                    periodo = request.session['periodo']
                    output = io.BytesIO()
                    workbook = xlsxwriter.Workbook(output)
                    worksheet = workbook.add_worksheet()
                    formatorojo = workbook.add_format({'bold': True, 'font_color': 'red'})
                    fil = 2
                    worksheet.write(fil, 0, '#')
                    worksheet.write(fil, 1, 'IDENTIFICACIÓN')
                    worksheet.write(fil, 2, 'NOMBRES Y APELLIDOS')
                    worksheet.write(fil, 3, 'CARRERA')
                    worksheet.write(fil, 4, 'PARALELO')
                    worksheet.write(fil, 5, 'FECHA')
                    worksheet.write(fil, 6, 'HORA INICIO')
                    worksheet.write(fil, 7, 'HORA FIN')
                    worksheet.write(fil, 8, 'PROFESOR1')
                    worksheet.write(fil, 9, 'PROFESOR2')
                    worksheet.write(fil, 10, 'PROFESOR3')
                    worksheet.write(fil, 11, 'PROFESOR4')
                    worksheet.write(fil, 12, 'GRUPO')
                    worksheet.write(fil, 13, 'SEDE')
                    worksheet.write(fil, 14, 'AULA')
                    worksheet.set_column(0, 0, 5)
                    worksheet.set_column(1, 1, 15)
                    worksheet.set_column(2, 2, 40)
                    worksheet.set_column(3, 3, 40)
                    worksheet.set_column(4, 4, 10)
                    worksheet.set_column(5, 5, 12)
                    worksheet.set_column(6, 6, 15)
                    worksheet.set_column(7, 7, 15)
                    worksheet.set_column(8, 8, 40)
                    worksheet.set_column(9, 9, 40)
                    worksheet.set_column(10, 10, 40)
                    worksheet.set_column(11, 11, 40)
                    worksheet.set_column(12, 12, 10)
                    worksheet.set_column(13, 13, 30)
                    worksheet.set_column(14, 14, 10)
                    fil = fil + 1
                    con = 1
                    listado1 = Materia.objects.filter(asignaturamalla__validarequisitograduacion=True,
                                                      nivel__periodo=periodo, asignaturamalla__status=True, status=True)
                    listado2 = Materia.objects.filter(asignaturamalla__malla_id=383, nivel__periodo=periodo,
                                                      asignaturamalla__status=True, status=True)
                    listado = listado1 | listado2
                    materiaasignada = MateriaAsignada.objects.filter(materia__in=listado)

                    for asignada in materiaasignada:
                        disermatasignada = DisertacionMateriaAsignadaPlanificacion.objects.filter(
                            materiaasignada=asignada)
                        if disermatasignada:
                            grupoplanificacion = DisertacionGrupoPlanificacion.objects.filter(pk=disermatasignada[0].grupoplanificacion.id).first()
                            tribunal = DisertacionTribunalPlanificacion.objects.filter(grupoplanificacion=grupoplanificacion).order_by('fecha_creacion')

                            worksheet.write(fil, 0, str(con))
                            worksheet.write(fil, 1,
                                            str(asignada.matricula.inscripcion.persona.identificacion()))
                            worksheet.write(fil, 2, str(asignada.matricula.inscripcion.persona))
                            worksheet.write(fil, 3, str(asignada.materia.asignaturamalla.malla.carrera))
                            worksheet.write(fil, 4, str(asignada.materia.paralelomateria.nombre))
                            worksheet.write(fil, 5,
                                            str(grupoplanificacion.aulaplanificacion.turnoplanificacion.fechaplanificacion.fecha))
                            worksheet.write(fil, 6, str(grupoplanificacion.aulaplanificacion.turnoplanificacion.horainicio))
                            worksheet.write(fil, 7, str(grupoplanificacion.aulaplanificacion.turnoplanificacion.horafin))

                            for i in range(4):
                                if i < len(tribunal):
                                    worksheet.write(fil, 8 + i, str(
                                        tribunal[i].responsable))
                                else:
                                    worksheet.write(fil, 8 + i, '')

                            worksheet.write(fil, 12, str(grupoplanificacion.grupo))
                            worksheet.write(fil, 13,
                                            str(grupoplanificacion.aulaplanificacion.turnoplanificacion.fechaplanificacion.sede))
                            worksheet.write(fil, 14, str(grupoplanificacion.aulaplanificacion.aula.nombre))
                        else:
                            worksheet.write(fil, 0, str(con))
                            worksheet.write(fil, 1,
                                            str(asignada.matricula.inscripcion.persona.identificacion()))
                            worksheet.write(fil, 2, str(asignada.matricula.inscripcion.persona))
                            worksheet.write(fil, 3, str(asignada.materia.asignaturamalla.malla.carrera))
                            worksheet.write(fil, 4, str(asignada.materia.paralelomateria.nombre))
                            worksheet.write(fil, 5,'No Asignado')
                            worksheet.write(fil, 6,'No Asignado')
                            worksheet.write(fil, 7,'No Asignado')
                            for i in range(4):
                                worksheet.write(fil, 8 + i, 'No Asignado')

                            worksheet.write(fil, 12, 'No Asignado')
                            worksheet.write(fil, 13,'No Asignado')
                            worksheet.write(fil, 14, 'No Asignado')

                        con += 1
                        fil += 1
                    workbook.close()
                    output.seek(0)
                    filename = 'reporte_disertaciones' + random.randint(1, 10000).__str__() + '.xlsx'
                    response = HttpResponse(output,
                                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    response['Content-Disposition'] = 'attachment; filename=%s' % filename
                    return response
                except Exception as ex:
                    pass

            elif action == 'adddisertacion':
                try:
                    data['titulo'] = u'Planificar disertacion'
                    form = DisertacionExamenComplexivoForm()
                    data['id'] = id = request.GET['idma']
                    data['cdla'] = request.GET['cdla']
                    asignada = MateriaAsignada.objects.filter(pk=int(encrypt(id))).first()
                    disermatasignada = DisertacionMateriaAsignadaPlanificacion.objects.filter(
                        materiaasignada=asignada).first()
                    if disermatasignada:
                        grupoplanificacion = DisertacionGrupoPlanificacion.objects.filter(
                            pk=disermatasignada.grupoplanificacion.id).first()
                        tribunal = DisertacionTribunalPlanificacion.objects.filter(
                            grupoplanificacion=grupoplanificacion).order_by('fecha_creacion')
                        aulaplanificacion = DisertacionAulaPlanificacion.objects.get(
                            pk=grupoplanificacion.aulaplanificacion.id)
                        turnoplanificacion = DisertacionTurnoPlanificacion.objects.get(
                            pk=aulaplanificacion.turnoplanificacion.id)
                        fechaplanificacion = DisertacionFechaPlanificacion.objects.get(
                            pk=turnoplanificacion.fechaplanificacion.id)
                        profesores = [
                            tribunal[i].responsable if len(tribunal) > i and tribunal[i].responsable else None
                            for i in range(4)
                        ]
                        responsable = grupoplanificacion.responsable
                        fecha_formateada = fechaplanificacion.fecha.strftime('%d-%m-%Y')
                        form = DisertacionExamenComplexivoForm(initial={
                            'fecha': fecha_formateada,
                            'horainicio': grupoplanificacion.aulaplanificacion.turnoplanificacion.horainicio,
                            'horafin': grupoplanificacion.aulaplanificacion.turnoplanificacion.horafin,
                            'profesor1':  profesores[0],
                            'profesor2':  profesores[1],
                            'profesor3':  profesores[2],
                            'profesor4':  profesores[3],
                            'sede':  grupoplanificacion.aulaplanificacion.turnoplanificacion.fechaplanificacion.sede.id,
                            'aula': grupoplanificacion.aulaplanificacion.aula.id,
                            'grupo': grupoplanificacion.grupo })
                        form.edit(profesores)
                        data['action'] = 'aditdisertacionalumno'
                    else:
                        data['action'] = 'adddisertacionalumno'
                    data['form'] = form
                    template = get_template('adm_complexivotematica/modal/formdisertacion.html')
                    return JsonResponse({'result': "ok", 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({'result': "bad", 'mensaje': f'Error: {ex}'})

            elif action == 'visualizacionhorario':
                try:
                    asignada = MateriaAsignada.objects.filter(materia_id=int(encrypt(request.GET['id'])))

                    for matasignada in asignada:
                        matasignada.visiblehorarioexamen = True
                        matasignada.save()
                    return JsonResponse({"result": "ok","mensaje": u"Horario de examen activado."})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

            elif action == 'desactivarvisualizacionhorario':
                try:
                    asignada = MateriaAsignada.objects.filter(materia_id=int(encrypt(request.GET['id'])))

                    for matasignada in asignada:
                        matasignada.visiblehorarioexamen = False
                        matasignada.save()
                    return JsonResponse({"result": "ok","mensaje": u"Horario de examen desactivado."})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

            return HttpResponseRedirect(request.path)
        else:
            try:
                data['title'] = u"Cupos de trabajo de titulación"
                try:
                    puede_realizar_accion(request, 'sga.puede_editar_cupotematica')
                    cordinacion = Coordinacion.objects.filter(status=True).exclude(pk__gte=6).order_by('-id')
                except Exception as ex:
                    cordinacion = persona.mis_coordinaciones().exclude(pk__gte=6).order_by('-id')
                data['facultades'] = cordinacion
                facultad = int(request.GET['fac']) if 'fac' in request.GET else 0
                carrera = int(request.GET['car']) if 'car' in request.GET else 0
                opci = int(request.GET['opc']) if 'opc' in request.GET else 0
                periodotitulacion = int(request.GET['per']) if 'per' in request.GET else 0
                search = None
                data['titperiodos'] = periodos = PeriodoGrupoTitulacion.objects.filter(status=True).order_by('-id')
                data['perid'] = periodos[0]
                tematicas = ComplexivoTematica.objects.filter(status=True, tutor__status=True).order_by('-tematica__vigente', 'tematica')
                if 's' in request.GET:
                    search = request.GET['s'].strip()
                    ss = search.split(' ')
                    if len(ss) == 1:
                        tematicas = tematicas.filter(Q(status=True) & (Q(tematica__tema__icontains=search) | Q(tutor__participante__persona__nombres__icontains=search) | Q(tutor__participante__persona__apellido1__icontains=search) | Q(tutor__participante__persona__apellido2__icontains=search)))
                    else:
                        tematicas = tematicas.filter(Q(status=True) & ((Q(tematica__tema__icontains=ss[0]) & Q(tematica__tema__icontains=ss[1])) | (Q(tutor__participante__persona__nombres__icontains=ss[0]) & Q(tutor__participante__persona__nombres__icontains=ss[1])) | (Q(tutor__participante__persona__apellido1__icontains=ss[0]) & Q(tutor__participante__persona__apellido2__icontains=ss[1]))))
                    if 'per' in request.GET:
                        data['perid'] = PeriodoGrupoTitulacion.objects.get(status=True, pk=int(request.GET['per']))
                        tematicas = tematicas.filter(periodo_id=request.GET['per'])
                    if 'fac' in request.GET:
                        data['facuid'] = cordinacion = Coordinacion.objects.get(pk=facultad)
                        data['carreras'] = cordinacion.carrera.filter(status=True)
                        tematicas = tematicas.filter(carrera__coordinacion=int(request.GET['fac']))
                    if 'car' in request.GET:
                        data['facuid'] = cordinacion = Coordinacion.objects.get(pk=facultad)
                        data['car_id'] = Carrera.objects.get(pk=int(request.GET['car']), status=True)
                        data['carreras'] = cordinacion.carrera.filter(status=True)
                        tematicas = tematicas.filter(carrera_id=request.GET['car'])
                    if 'opc' in request.GET:
                        data['op'] = int(request.GET['opc'])
                        if int(request.GET['opc']) == 1:
                            tematicas = tematicas.filter(complexivogrupotematica__isnull=False)
                        if int(request.GET['opc']) == 2:
                            tematicas = tematicas.filter(complexivogrupotematica__isnull=True)
                else:
                    if periodotitulacion > 0:
                        data['perid'] = PeriodoGrupoTitulacion.objects.get(pk=periodotitulacion)
                        tematicas = tematicas.filter(periodo_id=periodotitulacion)
                    else:
                        data['perid'] = periodos[0]
                        tematicas = tematicas.filter(periodo=periodos[0])
                    if facultad > 0:
                        tematicas = tematicas.filter(carrera__coordinacion=facultad)
                        data['facuid'] = cordinacion = Coordinacion.objects.get(pk=facultad)

                        data['carreras'] = car = cordinacion.carrera.filter(status=True)

                    if carrera > 0:
                        tematicas = tematicas.filter(carrera=carrera)
                        data['car_id'] = Carrera.objects.get(pk=carrera)
                    sds = []
                    listatematicas = []
                    data['op'] = opci
                    if opci > 0:
                        for tem in tematicas:
                            if opci == 1:
                                if tem.tiene_grupos():
                                    sds.append(tem)
                            else:
                                if not tem.tiene_grupos():
                                    sds.append(tem)
                        tematicas = sds

                if 'id' in request.GET:
                    data['id'] = id = int(request.GET['id'])
                    tematicas = tematicas.filter(id=id)
                tematicas = tematicas.order_by("tematica", "tutor__participante__persona__apellido1", "tutor__participante__persona__apellido2", "tutor__participante__persona__nombres")

                paging = MiPaginador(tematicas, 20)
                p = 1
                try:
                    paginasesion = 1
                    if 'paginador' in request.session:
                        paginasesion = int(request.session['paginador'])
                    if 'page' in request.GET:
                        p = int(request.GET['page'])
                    else:
                        p = paginasesion
                    try:
                        page = paging.page(p)
                    except:
                        p = 1
                    page = paging.page(p)
                except:
                    page = paging.page(p)
                request.session['paginador'] = p
                data['facultad'] = facultad
                data['carrera'] = carrera
                data['paging'] = paging
                data['rangospaging'] = paging.rangos_paginado(p)
                data['page'] = page
                data['tematicas'] = page.object_list
                data['search'] = search if search else ""

                # ---------------------------
                return render(request, "adm_complexivotematica/tematicacupo.html", data)
            except Exception as ex:
                return HttpResponseRedirect("/")

def valida_matricular_estudiante(data, alter, inscripcion):
    vali_alter = 0
    vali_tenido = 0
    data['item'] = alter
    data['grupo'] = alter.grupotitulacion
    data['inscripcion'] = inscripcion
    malla = inscripcion.inscripcionmalla_set.filter(status=True)[0].malla
    perfil = inscripcion.persona.mi_perfil()
    data['tiene_discapidad'] = perfil.tienediscapacidad
    if alter.estadofichaestudiantil:
        vali_alter += 1
        ficha = 0
        if inscripcion.persona.nombres and inscripcion.persona.apellido1 and inscripcion.persona.apellido2 and inscripcion.persona.nacimiento and inscripcion.persona.cedula and inscripcion.persona.nacionalidad and inscripcion.persona.email and inscripcion.persona.estado_civil and inscripcion.persona.sexo:
            data['datospersonales'] = True
            ficha += 1
        if inscripcion.persona.paisnacimiento and inscripcion.persona.provincianacimiento and inscripcion.persona.cantonnacimiento and inscripcion.persona.parroquianacimiento:
            data['datosnacimientos'] = True
            ficha += 1
        examenfisico = inscripcion.persona.datos_examen_fisico()
        if inscripcion.persona.sangre and examenfisico.peso and examenfisico.talla:
            data['datosmedicos'] = True
            ficha += 1
        if inscripcion.persona.pais and inscripcion.persona.provincia and inscripcion.persona.canton and inscripcion.persona.parroquia and inscripcion.persona.direccion and inscripcion.persona.direccion2 and inscripcion.persona.num_direccion and inscripcion.persona.telefono_conv or inscripcion.persona.telefono:
            data['datosdomicilio'] = True
            ficha += 1
        if perfil.raza:
            data['etnia'] = True
            ficha += 1
        if ficha == 5:
            vali_tenido += 1
    if alter.estadopracticaspreprofesionales:
        vali_alter += 1
        totalhoras = 0
        practicaspreprofesionalesinscripcion = PracticasPreprofesionalesInscripcion.objects.filter(inscripcion=inscripcion, status=True, culminada=True)
        data['malla_horas_practicas'] = malla.horas_practicas
        if practicaspreprofesionalesinscripcion.exists():
            for practicas in practicaspreprofesionalesinscripcion:
                if practicas.tiposolicitud == 3:
                    totalhoras += practicas.horahomologacion if practicas.horahomologacion else 0
                else:
                    totalhoras += practicas.numerohora
            if totalhoras >= malla.horas_practicas:
                data['practicaspreprofesionales'] = True
                vali_tenido += 1
        data['practicaspreprofesionalesvalor'] = totalhoras
    if alter.estadocredito:
        vali_alter += 1
        data['creditos'] = inscripcion.aprobo_asta_penultimo_malla()
        if inscripcion.aprobo_asta_penultimo_malla() and inscripcion.esta_matriculado_ultimo_nivel():
            vali_tenido += 1
        data['cantasigaprobadas'] = inscripcion.cantidad_asig_aprobada_penultimo_malla()
        data['cantasigaprobar'] = inscripcion.cantidad_asig_aprobar_penultimo_malla()
        data['esta_mat_ultimo_nivel'] = inscripcion.esta_matriculado_ultimo_nivel()
    if alter.estadoadeudar:
        vali_alter += 1
        if inscripcion.adeuda_a_la_fecha() == 0:
            data['deudas'] = True
            vali_tenido += 1
        data['deudasvalor'] = inscripcion.adeuda_a_la_fecha()
    if alter.estadoingles:
        vali_alter += 1
        modulo_ingles = ModuloMalla.objects.filter(malla=malla, status=True).exclude(asignatura_id=782)
        numero_modulo_ingles = modulo_ingles.count()
        lista = []
        listaid = []
        for modulo in modulo_ingles:
            if inscripcion.estado_asignatura(modulo.asignatura) == NOTA_ESTADO_APROBADO:
                lista.append(modulo.asignatura.nombre)
                listaid.append(modulo.asignatura.id)
        data['modulo_ingles_aprobados'] = lista
        data['modulo_ingles_faltante'] = modulo_ingles.exclude(asignatura_id__in=[int(i) for i in listaid])
        if numero_modulo_ingles == len(listaid):
            data['modulo_ingles'] = True
            vali_tenido += 1
    if alter.estadonivel:
        vali_alter += 1
        total_materias_malla = malla.cantidad_materiasaprobadas()
        cantidad_materias_aprobadas_record = inscripcion.recordacademico_set.filter(aprobada=True, status=True, asignatura__in=[x.asignatura for x in malla.asignaturamalla_set.filter(status=True)]).count()
        poraprobacion = round(cantidad_materias_aprobadas_record * 100 / total_materias_malla, 0)
        data['mi_nivel'] = nivel = inscripcion.mi_nivel()
        inscripcionmalla = inscripcion.malla_inscripcion()
        niveles_maximos = inscripcionmalla.malla.niveles_regulares
        # MOMENTANEO EL AUMENTO DE VALI_TENIDO
        vali_tenido += 1
        if poraprobacion >= 100:
            data['nivel'] = True
            # vali_tenido += 1
        else:
            if niveles_maximos == nivel.nivel.id:
                data['septimo'] = True
    if perfil.tienediscapacidad:
        data['discapacidad'] = perfil
    if inscripcion.persona.sexo.id == ESTADO_GESTACION:
        data['femenino'] = True
    if alter.estadovinculacion:
        vali_alter += 1
        data['malla_horas_vinculacion'] = malla.horas_vinculacion
        horastotal = ParticipantesMatrices.objects.filter(matrizevidencia_id=2, status=True, proyecto__status=True, inscripcion_id=inscripcion.id).aggregate(horastotal=Sum('horas'))['horastotal']
        horastotal = horastotal if horastotal else 0
        if horastotal >= malla.horas_vinculacion:
            data['vinculacion'] = True
            vali_tenido += 1
        data['horas_vinculacion'] = horastotal
    if alter.estadocomputacion:
        vali_alter += 1
        asignatura = AsignaturaMalla.objects.values_list('asignatura__id', flat=True).filter(malla__id=32)
        data['record_computacion'] = record = RecordAcademico.objects.filter(inscripcion__id=inscripcion.id, asignatura__id__in=asignatura, aprobada=True)
        creditos_computacion = 0
        data['malla_creditos_computacion'] = malla.creditos_computacion
        for comp in record:
            creditos_computacion += comp.creditos
        if creditos_computacion >= malla.creditos_computacion:
            data['computacion'] = True
            vali_tenido += 1
        data['creditos_computacion'] = creditos_computacion
    if vali_alter == vali_tenido:
        data['aprueba'] = True
    if inscripcion.persona.tipocelular == 0:
        data['tipocelular'] = '-'
    else:
        data['tipocelular'] = TIPO_CELULAR[int(inscripcion.persona.tipocelular) - 1][1]
    return data


def adicionar_nota_complexivo(idgraduado, nota, fecha, request):
    if ExamenComlexivoGraduados.objects.filter(graduado_id=idgraduado, itemexamencomplexivo_id=2).exists():
        itendetalle = ExamenComlexivoGraduados.objects.get(graduado_id=idgraduado, itemexamencomplexivo_id=2)
        itendetalle.examen = nota
        itendetalle.ponderacion = null_to_decimal((nota / 2), 2)
        itendetalle.fecha = fecha
        log(u'Adicionó Examen Complexivo graduado: %s' % itendetalle, request, "edit")
    else:
        itendetalle = ExamenComlexivoGraduados(graduado_id=idgraduado,
                                               itemexamencomplexivo_id=2,
                                               examen=nota,
                                               ponderacion=null_to_decimal((nota / 2), 2),
                                               fecha=fecha
                                               )
        log(u'Adicionó Examen Complexivo graduado por tribunal: %s' % itendetalle, request, "add")
    itendetalle.save(request)


def adicionoagrupofirma(request, lmate):
    tiporubrica = 3
    if lmate.asignaturamalla.malla.carrera.id == 208:
        if GrupoTitulacionIC.objects.values('id').filter(materia=lmate).exists():
            grupomateria = GrupoTitulacionIC.objects.filter(materia=lmate)[0]
            if grupomateria.tiporubrica:
                tiporubrica = grupomateria.tiporubrica
            else:
                tiporubrica = 3
        else:
            tiporubrica = 3
    if pertenecepredecesoratitulacion(lmate.asignaturamalla.id):
        tiporubrica = 2
    if predecesoratitulacion(lmate.asignaturamalla.id):
        tiporubrica = 4
    grupo = GrupoTitulacionIC.objects.filter(materia=lmate, tiporubrica=tiporubrica).first()
    if not grupo:
        grupo = GrupoTitulacionIC(materia=lmate, tiporubrica=tiporubrica)
        grupo.save(request)
        if not grupo.grupofirma_set.values('id').filter(status=True).exists():
            if tiporubrica in [3, 4]:
                grupofirma = GrupoFirma(grupo_id=grupo.id, activo=True)
                grupofirma.save(request)
                if not grupofirma.firmagrupotitulacion_set.values('id').filter(status=True).exists():
                    if lmate.profesor_materia_principal():
                        firmagrupo = FirmaGrupoTitulacion(grupofirma_id=grupofirma.id,
                                                          profesor_id=lmate.profesor_materia_principal().profesor.id)
                        firmagrupo.save(request)

    else:
        grupofirma = GrupoFirma.objects.filter(grupo_id=grupo.id, activo=True).first()
        if not grupofirma:
            if tiporubrica in [3, 4]:
                grupofirma = GrupoFirma(grupo_id=grupo.id, activo=True)
                grupofirma.save(request)
                if not grupofirma.firmagrupotitulacion_set.values('id').filter(status=True).exists():
                    if lmate.profesor_materia_principal():
                        firmagrupo = FirmaGrupoTitulacion(grupofirma_id=grupofirma.id,
                                                          profesor_id=lmate.profesor_materia_principal().profesor.id)
                        firmagrupo.save(request)
        else:
            if not grupofirma.firmagrupotitulacion_set.values('id').filter(status=True).exists():
                if lmate.profesor_materia_principal():
                    firmagrupo = FirmaGrupoTitulacion(grupofirma_id=grupofirma.id,
                                                      profesor_id=lmate.profesor_materia_principal().profesor.id)
                    firmagrupo.save(request)