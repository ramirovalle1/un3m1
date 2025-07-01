import json
import random
# decoradores
from time import sleep

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.template.loader import get_template
from django.forms import model_to_dict, TimeInput
from decorators import last_access, secure_module

from django.template import Context
from django.db import transaction
from django.http import HttpResponseRedirect, JsonResponse, HttpResponse
from django.shortcuts import render
from django.db.models.query_utils import Q
from datetime import datetime,date

from postulate.models import DEDICACION_PARTIDA, MODALIDAD_PARTIDA, RequisitosConvocatoriaPostulate, PeriodoConvocatoria, TipoPersonaConvocatoria, GrupoRequisitoConvocatoria, GrupoConvocatoria, PersonaPeriodoConvocatoria, PersonaRequisitosConvocatoria, ESTADO_POSTULANTE_CONVOCATORIA, HistorialPersonaPeriodoConvocatoria
from postulate.postular import validar_campos
from settings import EMAIL_INSTITUCIONAL_AUTOMATICO, ACTUALIZAR_FOTO_ALUMNOS
from sga.commonviews import adduserdata
from postulate.forms import PeriodoConvocatoriaForm, RequisitosConvocatoriaPostulateForm, TipoPersonaConvocatoriaForm, \
    GrupoRequisitoConvocatoriaForm, GrupoConvocatoriaForm, PostulanteMasivoForm, ValidarRequisitoForm, \
    NotificacionMasivaForm, PersonaPeriodoConvocatoriaForm
from sga.funciones import log, MiPaginador, numero_a_letras, remover_caracteres_especiales_unicode, generar_nombre,remover_caracteres_tildes_unicode
from sga.funcionesxhtml2pdf import conviert_html_to_pdf
from sagest.models import DenominacionPuesto
from sga.models import Carrera, Coordinacion, CUENTAS_CORREOS
from sga.models import AreaConocimientoTitulacion, SubAreaConocimientoTitulacion, SubAreaEspecificaConocimientoTitulacion, Carrera, Asignatura, Titulo, Persona
from sga.tasks import send_html_mail
from sga.templatetags.sga_extras import encrypt
from xlwt import *
import xlwt
import xlsxwriter
import io
import openpyxl

@login_required(redirect_field_name='ret', login_url='/loginpostulate')
# @secure_module
# @last_access
@transaction.atomic()
def view(request):
    global ex
    data = {}
    perfilprincipal = request.session['perfilprincipal']
    persona = request.session['persona']
    data['hoy'] = hoy = datetime.now().date()
    data['currenttime'] = datetime.now()
    data['perfil'] = persona.mi_perfil()

    if request.method == 'POST':
        action = request.POST['action']

        if action == 'addperiodo':
            try:
                form = PeriodoConvocatoriaForm(request.POST)
                if form.is_valid():
                    filtro = PeriodoConvocatoria(periodoacademico=form.cleaned_data['periodoacademico'],
                                                       descripcion=form.cleaned_data['descripcion'],
                                                       grupo=form.cleaned_data['grupo'],
                                                       requisitos=form.cleaned_data['requisitos'],
                                                       finicio=form.cleaned_data['finicio'],
                                                       ffin=form.cleaned_data['ffin'],
                                                       vigente=form.cleaned_data['vigente'])
                    filtro.save(request)
                    log(u'Agrego de Periodo Convocatoria: %s' % filtro, request, "addperiodo")
                    return JsonResponse({'result': False, 'mensaje': u'Guardado con exito'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': u'Error al guardar los datos'})

        if action == 'editperiodo':
            try:
                form = PeriodoConvocatoriaForm(request.POST)
                filtro = PeriodoConvocatoria.objects.get(id=int(encrypt(request.POST['id'])))
                if form.is_valid():
                    filtro.periodoacademico = form.cleaned_data['periodoacademico']
                    filtro.descripcion = form.cleaned_data['descripcion']
                    filtro.grupo = form.cleaned_data['grupo']
                    filtro.requisitos = form.cleaned_data['requisitos']
                    filtro.finicio = form.cleaned_data['finicio']
                    filtro.ffin = form.cleaned_data['ffin']
                    filtro.vigente = form.cleaned_data['vigente']
                    filtro.save(request)
                    log(u'Edicion de Periodo Convocatoria: %s' % filtro, request, "editperiodo")
                    return JsonResponse({'result': False, 'mensaje': 'Edicion Exitosa'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': u'Error al guardar los datos'})

        if action == 'vigenteperiodo':
            try:
                filtro = PeriodoConvocatoria.objects.get(id=int(encrypt(request.POST['id'])))
                filtro.vigente = request.POST['vigente']
                filtro.save(request)
                log(u'Edicion de estado vigencia de Periodo de Convocatoria: %s' % filtro, request, "vigenteperiodo")
                return JsonResponse({'result': 'ok', 'mensaje': 'Edicion Exitosa'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', 'mensaje': u'Error al guardar los datos'})

        if action == 'delperiodo':
            try:
                with transaction.atomic():
                    filtro = PeriodoConvocatoria.objects.get(id=int(encrypt(request.POST['id'])))
                    filtro.status = False
                    filtro.save(request)
                    log(u'Eliminacion de Periodo Convocatoria: %s' % filtro, request, "delperiodo")
                    res_json = {"error": False}
            except Exception as ex:
                res_json = {'error': True, "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)

        if action == 'addpostulantemasivo':
            try:
                form = PostulanteMasivoForm(request.POST, request.FILES)
                cab = PeriodoConvocatoria.objects.get(id=int(encrypt(request.POST['id'])))
                requisitos_ = cab.requisitos.traerrequisitos()
                if form.is_valid():
                    archivo = request.FILES['archivo']
                    wb = openpyxl.load_workbook(filename=archivo)
                    ws = wb.get_sheet_by_name("Hoja1")
                    cadena = ''
                    linea, excluidos, conexito = 0, 0, 0
                    worksheet = ws
                    lis_excluidos = []
                    linea_archivo = 1
                    for row in worksheet.iter_rows(min_row=0):
                        # if linea > 1:
                        if True:
                            currentValues, cadena = [], ''
                            for cell in row:
                                cadena += str(cell.value) + ' '
                                currentValues.append(str(cell.value))
                            carrera = currentValues[3]
                            modalidad = currentValues[5]
                            cedula = currentValues[6]
                            cargo = currentValues[8]
                            dedicacion = currentValues[9]
                            nivel = currentValues[10]
                            qspersona = Persona.objects.filter(status=True, cedula__icontains=cedula)
                            if qspersona.exists():
                                persona_ = qspersona.first()
                                if not PersonaPeriodoConvocatoria.objects.filter(status=True, periodo=cab, persona=persona_).exists():
                                    postulante = PersonaPeriodoConvocatoria(periodo=cab, persona=persona_, estado=0)
                                    dedicacion_ = DenominacionPuesto.objects.filter(status=True, descripcion__icontains=f'{cargo}').first()
                                    carrera_ = Carrera.objects.filter(status=True, nombre__icontains=f'{carrera}').first()
                                    coordinacion_ = carrera_.coordinaciones().first() if carrera_ else None
                                    if coordinacion_:
                                        postulante.coordinacion = coordinacion_
                                    if carrera_:
                                        postulante.carrera = carrera_
                                    if dedicacion_:
                                        postulante.denominacionpuesto = dedicacion_
                                    if modalidad and not modalidad == 'None':
                                        postulante.modalidad = [tup[0] for tup in MODALIDAD_PARTIDA if tup[1] == modalidad][0]
                                    if dedicacion:
                                        dedicacion_ = 1
                                        if dedicacion == 'TIEMPO COMPLETO':
                                            dedicacion_ = 1
                                        elif dedicacion == 'MEDIO TIEMPO':
                                            dedicacion_ = 2
                                        elif dedicacion == 'TIEMPO PARCIAL':
                                            dedicacion_ = 3
                                        postulante.dedicacion = dedicacion_
                                    postulante.save(request)
                                    for rl in requisitos_:
                                        for req in rl.requisito.all():
                                            requisito_ = PersonaRequisitosConvocatoria(tipo=rl.tipo, requisito=req, estado=0, participacion=postulante)
                                            requisito_.save(request)
                                    conexito += 1
                                else:
                                    excluidos += 1
                                    lis_excluidos.append(cedula)
                            else:
                                excluidos += 1
                                lis_excluidos.append(cedula)
                        linea += 1
                    messages.success(request, f'Total Cargados {conexito}')
                    log(u'Carga masiva de postulantes: %s' % cab, request, "addpostulantemasivo")
                    return JsonResponse({'result': False, 'mensaje': 'Edicion Exitosa'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': f'Error al guardar los datos. {ex}'})

        if action == 'actualizarpostulantemasivo':
            try:
                form = PostulanteMasivoForm(request.POST, request.FILES)
                cab = PeriodoConvocatoria.objects.get(id=int(encrypt(request.POST['id'])))
                if form.is_valid():
                    archivo = request.FILES['archivo']
                    wb = openpyxl.load_workbook(filename=archivo)
                    ws = wb.get_sheet_by_name("Hoja1")
                    cadena = ''
                    linea, excluidos, conexito = 0, 0, 0
                    worksheet = ws
                    lis_excluidos = []
                    linea_archivo = 1
                    for row in worksheet.iter_rows(min_row=0):
                        # if linea > 1:
                        if True:
                            currentValues, cadena = [], ''
                            for cell in row:
                                cadena += str(cell.value) + ' '
                                currentValues.append(str(cell.value))
                            facultad = currentValues[2]
                            cedula = currentValues[6]
                            qspersona = Persona.objects.filter(status=True, cedula__icontains=cedula)
                            if qspersona.exists():
                                persona_ = qspersona.first()
                                if PersonaPeriodoConvocatoria.objects.filter(status=True, periodo=cab, persona=persona_).exists():
                                    postulante = PersonaPeriodoConvocatoria.objects.filter(status=True, periodo=cab, persona=persona_).order_by('-id').first()
                                    coordinacion_ = Coordinacion.objects.filter(status=True,alias__icontains=f'{facultad}')
                                    if coordinacion_:
                                        postulante.coordinacion = coordinacion_.order_by('-id').first()
                                    postulante.save(request)
                                    conexito += 1
                                else:
                                    excluidos += 1
                                    lis_excluidos.append(cedula)
                            else:
                                excluidos += 1
                                lis_excluidos.append(cedula)
                        linea += 1
                    messages.success(request, f'Total actualizados {conexito}')
                    log(u'Actualizacion de facultad masiva de postulantes: %s' % cab, request, "aeditpostulantemasivo")
                    return JsonResponse({'result': False, 'mensaje': 'Edicion Exitosa'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': f'Error al guardar los datos. {ex}'})

        if action == 'validarrequisitos':
            try:
                form = ValidarRequisitoForm(request.POST)
                filtro = PersonaPeriodoConvocatoria.objects.get(id=int(encrypt(request.POST['id'])))
                if form.is_valid():
                    filtro.observacion_validacion = form.cleaned_data['descripcion']
                    filtro.fecha_validacion = datetime.now().date()
                    filtro.estado = form.cleaned_data['estado']
                    filtro.save(request)
                    historial = HistorialPersonaPeriodoConvocatoria(
                        personaperiodo = filtro,
                        observacion = form.cleaned_data['descripcion'],
                        estado = form.cleaned_data['estado']
                    )
                    historial.save(request)
                    send_html_mail("Validacion de requisitos", "emails/validacionrequisito.html",
                                   {'sistema': request.session['nombresistema'], 'fecha': datetime.now().date(), 'hora': datetime.now().time(),'filtro':filtro}, [filtro.persona.email,], [],
                                   cuenta=CUENTAS_CORREOS[30][1])
                    log(u'Valido Requisitos de Ingreso: %s' % filtro, request, "validarrequisitos")
                    return JsonResponse({'result': False, 'mensaje': 'Edicion Exitosa'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': u'Error al guardar los datos'})

        if action == 'delpostulante':
            try:
                with transaction.atomic():
                    filtro = PersonaPeriodoConvocatoria.objects.get(id=int(encrypt(request.POST['id'])))
                    filtro.status = False
                    filtro.save(request)
                    log(u'Eliminacion de Postulante Periodo Convocatoria: %s' % filtro, request, "delperiodo")
                    res_json = {"error": False}
            except Exception as ex:
                res_json = {'error': True, "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)

        if action == 'editpostulante':
            try:
                filtro = PersonaPeriodoConvocatoria.objects.get(id=int(encrypt(request.POST['id'])))
                filtro.carrera_id = request.POST['carrera']
                filtro.coordinacion = Carrera.objects.get(id=int(request.POST['carrera'])).coordinacion_set.first()
                filtro.modalidad = request.POST['modalidad']
                filtro.dedicacion = request.POST['dedicacion']
                filtro.save(request)
                log(u'Editó postulante: %s' % filtro.persona, request, "edit")
                return JsonResponse({'result': False, 'mensaje': 'Edicion Exitosa'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': f'Error al guardar los datos. {ex}'})

        if action == 'addrequisito':
            try:
                form = RequisitosConvocatoriaPostulateForm(request.POST,request.FILES)
                if form.is_valid():
                    filtro = RequisitosConvocatoriaPostulate(titulo=form.cleaned_data['titulo'],
                                                          descripcion=form.cleaned_data['descripcion'])
                    filtro.varchivo = form.cleaned_data['varchivo']
                    filtro.vdescripcion = form.cleaned_data['vdescripcion']
                    if 'formato' in request.FILES:
                        newfile = request.FILES['formato']
                        newfile._name = remover_caracteres_especiales_unicode(remover_caracteres_tildes_unicode(generar_nombre(f"formato_{filtro.titulo}_", newfile._name)))
                        filtro.formato = newfile
                    filtro.save(request)
                    log(u'Agrego de Requisito Convocatoria: %s' % filtro, request, "addrequisito")
                    return JsonResponse({'result': False, 'mensaje': u'Guardado con exito'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': u'Error al guardar los datos'})

        if action == 'editrequisito':
            try:
                form = RequisitosConvocatoriaPostulateForm(request.POST)
                filtro = RequisitosConvocatoriaPostulate.objects.get(id=int(encrypt(request.POST['id'])))
                if form.is_valid():
                    filtro.titulo = form.cleaned_data['titulo']
                    filtro.descripcion = form.cleaned_data['descripcion']
                    filtro.varchivo = form.cleaned_data['varchivo']
                    filtro.vdescripcion = form.cleaned_data['vdescripcion']
                    if 'formato' in request.FILES:
                        newfile = request.FILES['formato']
                        newfile._name = remover_caracteres_especiales_unicode(remover_caracteres_tildes_unicode(generar_nombre(f"formato_{filtro.titulo}_", newfile._name)))
                        filtro.formato = newfile
                    filtro.save(request)
                    log(u'Edicion de Requisito Convocatoria: %s' % filtro, request, "editrequisito")
                    return JsonResponse({'result': False, 'mensaje': 'Edicion Exitosa'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': u'Error al guardar los datos'})

        if action == 'delrequisito':
            try:
                with transaction.atomic():
                    filtro = RequisitosConvocatoriaPostulate.objects.get(id=int(encrypt(request.POST['id'])))
                    filtro.status = False
                    filtro.save(request)
                    log(u'Eliminacion de Requisito Convocatoria: %s' % filtro, request, "delrequisito")
                    res_json = {"error": False}
            except Exception as ex:
                res_json = {'error': True, "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)

        if action == 'addtipopersona':
            try:
                form = TipoPersonaConvocatoriaForm(request.POST)
                if form.is_valid():
                    filtro = TipoPersonaConvocatoria(descripcion=form.cleaned_data['descripcion'])
                    filtro.save(request)
                    log(u'Agrego de Grupo Persona Convocatoria: %s' % filtro, request, "addtipopersona")
                    return JsonResponse({'result': False, 'mensaje': u'Guardado con exito'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': u'Error al guardar los datos'})

        if action == 'edittipopersona':
            try:
                form = TipoPersonaConvocatoriaForm(request.POST)
                filtro = TipoPersonaConvocatoria.objects.get(id=int(encrypt(request.POST['id'])))
                if form.is_valid():
                    filtro.descripcion = form.cleaned_data['descripcion']
                    filtro.save(request)
                    log(u'Edicion de Grupo Persona Convocatoria: %s' % filtro, request, "edittipopersona")
                    return JsonResponse({'result': False, 'mensaje': 'Edicion Exitosa'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': u'Error al guardar los datos'})

        if action == 'deltipopersona':
            try:
                with transaction.atomic():
                    filtro = TipoPersonaConvocatoria.objects.get(id=int(encrypt(request.POST['id'])))
                    filtro.status = False
                    filtro.save(request)
                    log(u'Eliminacion de Grupo Persona Convocatoria: %s' % filtro, request, "deltipopersona")
                    res_json = {"error": False}
            except Exception as ex:
                res_json = {'error': True, "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)

        if action == 'addgrupo':
            try:
                form = GrupoConvocatoriaForm(request.POST)
                if form.is_valid():
                    filtro = GrupoConvocatoria(grupo=form.cleaned_data['grupo'], version=form.cleaned_data['version'], vigente=form.cleaned_data['vigente'])
                    filtro.save(request)
                    log(u'Agrego de Grupo Convocatoria: %s' % filtro, request, "addgrupo")
                    return JsonResponse({'result': False, 'mensaje': u'Guardado con exito'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': u'Error al guardar los datos'})

        if action == 'editgrupo':
            try:
                form = GrupoConvocatoriaForm(request.POST)
                filtro = GrupoConvocatoria.objects.get(id=int(encrypt(request.POST['id'])))
                if form.is_valid():
                    # filtro.grupo = form.cleaned_data['grupo']
                    filtro.vigente = form.cleaned_data['vigente']
                    filtro.version = form.cleaned_data['version']
                    filtro.save(request)
                    log(u'Edicion de Grupo Convocatoria: %s' % filtro, request, "editgrupo")
                    return JsonResponse({'result': False, 'mensaje': 'Edicion Exitosa'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': u'Error al guardar los datos'})

        if action == 'delgrupo':
            try:
                with transaction.atomic():
                    filtro = GrupoConvocatoria.objects.get(id=int(encrypt(request.POST['id'])))
                    filtro.status = False
                    filtro.save(request)
                    log(u'Eliminacion de Grupo Convocatoria: %s' % filtro, request, "delgrupo")
                    res_json = {"error": False}
            except Exception as ex:
                res_json = {'error': True, "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)

        if action == 'addgruporequisito':
            try:
                cab = GrupoConvocatoria.objects.get(id=int(encrypt(request.POST['id'])))
                form = GrupoRequisitoConvocatoriaForm(request.POST)
                if form.is_valid():
                    if GrupoRequisitoConvocatoria.objects.filter(tipo=form.cleaned_data['tipo'], grupo=cab, status=True).exists():
                        return JsonResponse({'result': True, 'mensaje': f'Ya existe este tipo de requito en {cab}'})
                    filtro = GrupoRequisitoConvocatoria(tipo=form.cleaned_data['tipo'], grupo=cab)
                    filtro.save(request)
                    requisito_ = request.POST.getlist('requisito', '')
                    for i in requisito_:
                        filtro.requisito.add(i)
                    log(u'Agrego de Grupo Requisito Convocatoria: %s' % filtro, request, "addgruporequisito")
                    return JsonResponse({'result': False, 'mensaje': u'Guardado con exito'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': u'Error al guardar los datos'})

        if action == 'editgruporequisito':
            try:
                form = GrupoRequisitoConvocatoriaForm(request.POST)
                filtro = GrupoRequisitoConvocatoria.objects.get(id=int(encrypt(request.POST['id'])))
                if form.is_valid():
                    filtro.tipo = form.cleaned_data['tipo']
                    # filtro.grupo = form.cleaned_data['grupo']
                    filtro.requisito.clear()
                    requisito_ = request.POST.getlist('requisito', '')
                    for i in requisito_:
                        filtro.requisito.add(i)
                    filtro.save(request)
                    log(u'Edicion de Grupo Requisito Convocatoria: %s' % filtro, request, "editgruporequisito")
                    return JsonResponse({'result': False, 'mensaje': 'Edicion Exitosa'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': u'Error al guardar los datos'})

        if action == 'delgruporequisito':
            try:
                with transaction.atomic():
                    filtro = GrupoRequisitoConvocatoria.objects.get(id=int(encrypt(request.POST['id'])))
                    filtro.status = False
                    filtro.save(request)
                    log(u'Eliminacion de Grupo Requisito Convocatoria: %s' % filtro, request, "delgruporequisito")
                    res_json = {"error": False}
            except Exception as ex:
                res_json = {'error': True, "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)

        if action == 'notificarperiodo':
            try:
                filtro = PeriodoConvocatoria.objects.get(id=int(encrypt(request.POST['id'])))
                if not PersonaPeriodoConvocatoria.objects.filter(status=True, periodo=filtro,estado = 0).exists():
                    raise NameError('No existen postulantes en estado pendiente.')
                if not 'mensaje' in request.POST:
                    raise  NameError('Ingresar mensaje de notificacion.')
                if not 'titulo' in request.POST:
                    raise  NameError('Ingresar titulo de notificacion.')
                qsbase = PersonaPeriodoConvocatoria.objects.filter(status=True, periodo=filtro,estado = 0)
                mensaje = request.POST['mensaje']
                titulo = request.POST['titulo']
                for listado in qsbase:
                    sleep(2)
                    send_html_mail(f"{titulo}", "emails/cargar_requisitos.html",
                                   {'sistema': request.session['nombresistema'],'titulomensaje': titulo, 'fecha': datetime.now().date(),'mensaje':mensaje, 'hora': datetime.now().time(), 'filtro': listado}, [listado.persona.email,], [],
                                   cuenta=CUENTAS_CORREOS[30][1])
                return JsonResponse({'result': False, 'mensaje': 'Notificacion Exitosa'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': u'Error al guardar los datos. Detalle: %s'%(ex.__str__())})

        return JsonResponse({"result": "bad", "mensaje": u"Solicitud Incorrecta."})
    else:
        adduserdata(request, data)
        if 'action' in request.GET:
            data['action'] = action = request.GET['action']

            if action == 'addperiodo':
                try:
                    data['title'] = u'Adiccionar Periodo Convocatoria'
                    form = PeriodoConvocatoriaForm()
                    form.fields['requisitos'].queryset = GrupoConvocatoria.objects.none()
                    data['form'] = form
                    template = get_template("postulate/adm_periodoconvocatoria/modal/formperiodo.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            if action == 'editperiodo':
                try:
                    data['title'] = u'Editar Periodo Convocatoria'
                    data['filtro'] = filtro = PeriodoConvocatoria.objects.get(id=int(encrypt(request.GET['id'])))
                    initial = model_to_dict(filtro)
                    form = PeriodoConvocatoriaForm(initial=initial)
                    form.fields['requisitos'].queryset = GrupoConvocatoria.objects.filter(status=True, grupo__id=filtro.grupo.id)
                    form.fields['finicio'].initial = str(filtro.finicio)
                    form.fields['ffin'].initial = str(filtro.ffin)
                    data['form'] = form
                    template = get_template("postulate/adm_periodoconvocatoria/modal/formperiodo.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass
            if action == 'editpostulante':
                try:
                    data['title'] = u'Editar postulante'
                    data['action'] = action
                    data['filtro'] = filtro = PersonaPeriodoConvocatoria.objects.get(id=int(encrypt(request.GET['id'])))
                    form = PersonaPeriodoConvocatoriaForm()
                    form.editar(filtro)
                    data['form'] = form
                    template = get_template("postulate/adm_periodoconvocatoria/modal/formpostulante.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            if action == 'buscacarrera':
                q = (request.GET['q']).replace(",", "")
                filtro = Carrera.objects.filter(status=True, nombre__icontains=q).exclude(coordinacion=9)
                resp = [{'id': qs.pk, 'text': f"{qs.nombre}"} for qs in filtro]
                return HttpResponse(json.dumps({'status': True, 'results': resp}))
            if action == 'set_coordinacion':
                id = (request.GET['id'])
                filtro = Carrera.objects.get(id=id).coordinacion_set.first() if Carrera.objects.filter(id=id).exists() else None
                if filtro:
                    resp = {'id': filtro.pk, 'text': filtro.nombre}
                    return HttpResponse(json.dumps({'result': 'ok', 'data': resp}))
            if action == 'consultarrequisitios':
                id = (request.GET['id']).replace(",", "")
                filtro = GrupoConvocatoria.objects.filter(status=True, grupo__id=int(id), vigente=True)
                if 'search' in request.GET:
                    search = request.GET['search']
                    filtro = filtro.filter(grupo__descripcion__icontains=search)
                resp = [{'id': cr.pk, 'text': "{} v{}".format( cr.grupo.__str__(), cr.version)} for cr in filtro.order_by('grupo__descripcion')]
                return HttpResponse(json.dumps({'state': True, 'result': resp}))

            if action == 'postulante':
                try:
                    id = request.GET['id']
                    estado, search, url_vars, filtro = request.GET.get('estado', ''), request.GET.get('s', ''), '',Q(status=True)
                    data['cab'] = cab = PeriodoConvocatoria.objects.get(id=id)
                    data['title'] = f'Postulates para  {cab}'
                    qsbase = PersonaPeriodoConvocatoria.objects.filter(status=True, periodo=cab)
                    url_vars = f'&action={action}&id={id}'
                    if estado:
                        data['estado'] = int(estado)
                        url_vars += "&estado={}".format(estado)
                        filtro = filtro & (Q(estado=estado))
                    if search:
                        data['search'] = search
                        s = search.split(' ')
                        url_vars += "&s={}".format(search)
                        if len(s)>1:
                            filtro = filtro & (Q(persona__apellido1__icontains=s[0]) | Q(persona__apellido2__icontains=s[1]))
                        else:
                            filtro = filtro & (Q(persona__cedula__icontains=s[0]) | Q(persona__apellido1__icontains=s[0]))
                    qsbase = qsbase.filter(filtro).order_by('persona__apellido1')
                    paging = MiPaginador(qsbase, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data["url_vars"] = url_vars
                    data['listado'] = page.object_list
                    data['ESTADO_POSTULANTE_CONVOCATORIA'] = ESTADO_POSTULANTE_CONVOCATORIA
                    data['list_count'] = len(qsbase)
                    return render(request, "postulate/adm_periodoconvocatoria/viewpostulantes.html", data)
                except Exception as ex:
                    pass

            if action == 'addpostulantemasivo':
                try:
                    data['title'] = u'Adiccionar Periodo Convocatoria'
                    data['filtro'] = filtro = PeriodoConvocatoria.objects.get(id=int(request.GET['id']))
                    form = PostulanteMasivoForm()
                    data['form'] = form
                    template = get_template("postulate/adm_periodoconvocatoria/modal/formpostulantemasivo.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            if action == 'actualizarpostulantemasivo':
                try:
                    data['title'] = u'Actualizar facultades de postulantes'
                    data['filtro'] = filtro = PeriodoConvocatoria.objects.get(id=int(request.GET['id']))
                    form = PostulanteMasivoForm()
                    data['form'] = form
                    template = get_template("postulate/adm_periodoconvocatoria/modal/formpostulantemasivo.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            if action == 'validarrequisitos':
                try:
                    data['title'] = u'Validar Requisitos'
                    data['filtro'] = filtro = PersonaPeriodoConvocatoria.objects.get(id=int(encrypt(request.GET['id'])))
                    form = ValidarRequisitoForm()
                    data['form'] = form
                    template = get_template("postulate/adm_periodoconvocatoria/modal/validarrequisitospostulante.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            if action == 'verrequisitospostulante':
                try:
                    data['title'] = u'Ver Requisitos'
                    data['filtro'] = filtro = PersonaPeriodoConvocatoria.objects.get(id=int(encrypt(request.GET['id'])))
                    template = get_template("postulate/adm_periodoconvocatoria/modal/verrequisitospostulante.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'requisitos':
                try:
                    data['title'] = u'Requisitos Convocatoria'
                    search, url_vars, filtro = request.GET.get('s', ''), '',Q(status=True)
                    qsbase = RequisitosConvocatoriaPostulate.objects.filter(status=True)
                    url_vars = f'&action={action}'
                    if search:
                        data['search'] = search
                        url_vars += "&s={}".format(search)
                        filtro = filtro & (Q(titulo__icontains=search) | Q(descripcion__icontains=search))
                    qsbase = qsbase.filter(filtro).order_by('titulo')
                    paging = MiPaginador(qsbase, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data["url_vars"] = url_vars
                    data['listado'] = page.object_list
                    data['list_count'] = len(qsbase)
                    return render(request, "postulate/adm_periodoconvocatoria/viewrequisito.html", data)
                except Exception as ex:
                    pass

            elif action == 'addrequisito':
                try:
                    data['title'] = u'Adiccionar Convocatoria'
                    form = RequisitosConvocatoriaPostulateForm()
                    data['form'] = form
                    template = get_template("postulate/adm_periodoconvocatoria/modal/formrequisito.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'editrequisito':
                try:
                    data['title'] = u'Editar Convocatoria'
                    data['filtro'] = filtro = RequisitosConvocatoriaPostulate.objects.get(id = int(encrypt(request.GET['id'])))
                    initial = model_to_dict(filtro)
                    form = RequisitosConvocatoriaPostulateForm(initial=initial)
                    data['form'] = form
                    template = get_template("postulate/adm_periodoconvocatoria/modal/formrequisito.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'tipopersona':
                try:
                    data['title'] = u'Tipo Grupo Requisitos'
                    search, url_vars, filtro = request.GET.get('s', ''), '',Q(status=True)
                    qsbase = TipoPersonaConvocatoria.objects.filter(status=True)
                    url_vars = f'&action={action}'
                    if search:
                        data['search'] = search
                        url_vars += "&s={}".format(search)
                        filtro = filtro & (Q(descripcion__icontains=search))
                    qsbase = qsbase.filter(filtro).order_by('descripcion')
                    paging = MiPaginador(qsbase, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data["url_vars"] = url_vars
                    data['listado'] = page.object_list
                    data['list_count'] = len(qsbase)
                    return render(request, "postulate/adm_periodoconvocatoria/viewgrupo.html", data)
                except Exception as ex:
                    pass

            elif action == 'addtipopersona':
                try:
                    data['title'] = u'Adiccionar Tipo Grupo Convocatoria'
                    form = TipoPersonaConvocatoriaForm()
                    data['form'] = form
                    template = get_template("postulate/adm_periodoconvocatoria/modal/formtipo.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'edittipopersona':
                try:
                    data['title'] = u'Editar Tipo Grupo Convocatoria'
                    data['filtro'] = filtro = TipoPersonaConvocatoria.objects.get(id=int(encrypt(request.GET['id'])))
                    initial = model_to_dict(filtro)
                    form = TipoPersonaConvocatoriaForm(initial=initial)
                    data['form'] = form
                    template = get_template("postulate/adm_periodoconvocatoria/modal/formtipo.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'gruporequisito':
                try:
                    data['title'] = u'Grupo Requisito Convocatoria'
                    search, url_vars, filtro = request.GET.get('s', ''), '',Q(status=True)
                    qsbase = GrupoConvocatoria.objects.filter(status=True)
                    url_vars = f'&action={action}'
                    if search:
                        data['search'] = search
                        url_vars += "&s={}".format(search)
                        filtro = filtro & (Q(grupo__descripcion__icontains=search))
                    qsbase = qsbase.filter(filtro).order_by('grupo__descripcion', 'vigente')
                    paging = MiPaginador(qsbase, 4)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data["url_vars"] = url_vars
                    data['listado'] = page.object_list
                    data['list_count'] = len(qsbase)
                    return render(request, "postulate/adm_periodoconvocatoria/viewgruporequisito.html", data)
                except Exception as ex:
                    pass

            elif action == 'addgrupo':
                try:
                    data['title'] = u'Adiccionar Grupo Convocatoria'
                    form = GrupoConvocatoriaForm()
                    data['form'] = form
                    template = get_template("postulate/adm_periodoconvocatoria/modal/formgrupo.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'editgrupo':
                try:
                    data['title'] = u'Editar Grupo'
                    data['filtro'] = filtro = GrupoConvocatoria.objects.get(id=int(encrypt(request.GET['id'])))
                    initial = model_to_dict(filtro)
                    form_ = GrupoConvocatoriaForm(initial=initial)
                    del form_.fields['grupo']
                    data['form'] = form_
                    template = get_template("postulate/adm_periodoconvocatoria/modal/formgrupo.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'addgruporequisito':
                try:
                    data['title'] = u'Adiccionar Grupo Requisito Convocatoria'
                    data['filtro'] = filtro = GrupoConvocatoria.objects.get(id=int(request.GET['id']))
                    form = GrupoRequisitoConvocatoriaForm()
                    data['form'] = form
                    template = get_template("postulate/adm_periodoconvocatoria/modal/formgruporequisito.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'editgruporequisito':
                try:
                    data['title'] = u'Editar Grupo Requisito Convocatoria'
                    data['filtro'] = filtro = GrupoRequisitoConvocatoria.objects.get(id=int(encrypt(request.GET['id'])))
                    initial = model_to_dict(filtro)
                    form_ = GrupoRequisitoConvocatoriaForm(initial=initial)
                    data['form'] = form_
                    template = get_template("postulate/adm_periodoconvocatoria/modal/formgruporequisito.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'editgruporequisito':
                try:
                    data['title'] = u'Editar Grupo Requisito Convocatoria'
                    data['filtro'] = filtro = GrupoRequisitoConvocatoria.objects.get(id=int(encrypt(request.GET['id'])))
                    initial = model_to_dict(filtro)
                    form_ = GrupoRequisitoConvocatoriaForm(initial=initial)
                    data['form'] = form_
                    template = get_template("postulate/adm_periodoconvocatoria/modal/formgruporequisito.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'reportepostulantes':
                try:
                    id = request.GET['id']
                    __author__ = 'Unemi'
                    output = io.BytesIO()
                    workbook = xlsxwriter.Workbook(output)
                    ws = workbook.add_worksheet('postulantes')
                    formatoceldatitulo = workbook.add_format({'text_wrap': True, 'bg_color': 'silver', 'align': 'center'})
                    formatoceldacenter = workbook.add_format({'valign': 'vcenter'})
                    ws.set_column(0, 20, 50)
                    ws.write('A1', 'Fec. Creacion', formatoceldatitulo)
                    ws.write('B1', 'Postulante', formatoceldatitulo)
                    ws.write('C1', 'Cedula', formatoceldatitulo)
                    ws.write('D1', 'Telefono', formatoceldatitulo)
                    ws.write('E1', 'Correo', formatoceldatitulo)
                    ws.write('F1', 'Facultad', formatoceldatitulo)
                    ws.write('G1', 'Carrera', formatoceldatitulo)
                    ws.write('H1', 'Modalidad', formatoceldatitulo)
                    ws.write('I1', 'Jornada ', formatoceldatitulo)
                    ws.write('J1', 'Cargo ', formatoceldatitulo)
                    ws.write('K1', 'Dedicacion', formatoceldatitulo)
                    ws.write('L1', 'Estado', formatoceldatitulo)
                    cab = PeriodoConvocatoria.objects.get(id=id)
                    qsbase = PersonaPeriodoConvocatoria.objects.filter(status=True, periodo=cab)
                    requisitos = qsbase[0].traerrequisitos()
                    columna = 12
                    for list in range(len(requisitos)):
                        ws.write(0,columna,str(requisitos[list].requisito.titulo),formatoceldatitulo)
                        ws.write(0,columna+1,'Observación',formatoceldatitulo)
                        columna+=2

                    row_num = 1
                    for post in qsbase:
                        ws.write(row_num, 0, str(post.fecha_creacion.strftime("%Y-%m-%d %H:%M:%S")), formatoceldacenter)
                        ws.write(row_num, 1, str(post.persona.nombre_completo()), formatoceldacenter)
                        ws.write(row_num, 2, str(post.persona.cedula), formatoceldacenter)
                        ws.write(row_num, 3, str(post.persona.telefono), formatoceldacenter)
                        ws.write(row_num, 4, str(post.persona.email), formatoceldacenter)
                        ws.write(row_num, 5, str(post.coordinacion), formatoceldacenter)
                        ws.write(row_num, 6, str(post.carrera), formatoceldacenter)
                        ws.write(row_num, 7, str(post.get_dedicacion_display()), formatoceldacenter)
                        ws.write(row_num, 8, str(post.get_modalidad_display()), formatoceldacenter)
                        ws.write(row_num, 9, str(post.denominacionpuesto), formatoceldacenter)
                        ws.write(row_num, 10, str(post.get_dedicacion_display()), formatoceldacenter)
                        ws.write(row_num, 11, str(post.get_estado_display()), formatoceldacenter)
                        colum = 12
                        for lista in post.traerrequisitos():
                            ws.write(row_num,colum,'SI' if lista.estado >= 1 else 'NO',formatoceldacenter)
                            ws.write(row_num,colum+1,lista.observacion_revisor if lista else '',formatoceldacenter)
                            colum +=2
                        row_num += 1
                    workbook.close()
                    output.seek(0)
                    filename = 'reporte_postulantes' + random.randint(1, 10000).__str__() + '.xlsx'
                    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    response['Content-Disposition'] = 'attachment; filename=%s' % filename
                    return response
                except Exception as ex:
                    pass

            elif action == 'notificarperiodo':
                try:
                    data['title'] = u'Notificación'
                    data['filtro'] = filtro = PeriodoConvocatoria.objects.get(id=int(encrypt(request.GET['id'])))
                    form_ = NotificacionMasivaForm()
                    data['form'] = form_
                    template = get_template("postulate/adm_periodoconvocatoria/modal/formgruporequisito.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'verhistorial':
                try:
                    data['title'] = u'Ver Requisitos'
                    data['filtro'] = filtro = PersonaPeriodoConvocatoria.objects.get(id=int(encrypt(request.GET['id'])))
                    template = get_template("postulate/adm_periodoconvocatoria/modal/verhistorial.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass


        else:
            try:
                data['title'] = u'Periodo Requisitos de Ingresos'
                search, filtro, url_vars = request.GET.get('s', ''), (Q(status=True)), ''

                if search:
                    data['search'] = search
                    url_vars += "&s={}".format(search)
                    filtro = filtro & (Q(descripcion__icontains=search))

                listado = PeriodoConvocatoria.objects.filter(filtro).order_by('-id')
                paging = MiPaginador(listado, 20)
                p = 1
                try:
                    paginasesion = 1
                    if 'paginador' in request.session:
                        paginasesion = int(request.session['paginador'])
                    if 'page' in request.GET:
                        p = int(request.GET['page'])
                    else:
                        p = paginasesion
                    try:
                        page = paging.page(p)
                    except:
                        p = 1
                    page = paging.page(p)
                except:
                    page = paging.page(p)
                request.session['paginador'] = p
                data['paging'] = paging
                data['rangospaging'] = paging.rangos_paginado(p)
                data['page'] = page
                data["url_vars"] = url_vars
                data['listado'] = page.object_list
                data['list_count'] = len(listado)
                return render(request, "postulate/adm_periodoconvocatoria/view.html", data)
            except Exception as ex:
                pass

