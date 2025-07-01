#!/usr/bin/env python
# -*- coding: utf-8 -*-
import os
import sys
SITE_ROOT = os.path.dirname(os.path.realpath(__file__))
your_djangoproject_home = os.path.split(SITE_ROOT)[0]
sys.path.append(your_djangoproject_home)
from django.core.wsgi import get_wsgi_application
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "settings")
application = get_wsgi_application()
from django.http import HttpResponse
from sga.models import *
from sagest.models import *
from posgrado.models import *
from Moodle_Funciones import *

from openpyxl import load_workbook, workbook as openxl
from openpyxl.styles import Font as openxlFont
from openpyxl.styles.alignment import Alignment as alin
from datetime import datetime, timedelta, date
from settings import  MEDIA_ROOT
hoy = datetime.now().date()
url = ''
directory = os.path.join(SITE_STORAGE, 'media', 'auditoria')
try:
    os.stat(directory)
except:
    os.mkdir(directory)

# listadoperiodos = Periodo.objects.filter(pk__in=[110,112,113,119,126,153,177])
listadoperiodos = Periodo.objects.filter(pk__in=[177])
ARCHIVO = os.path.join(SITE_STORAGE, 'archivos')
totalperiodos = listadoperiodos.count()
cuentaperiodo = 0
for lper in listadoperiodos:
    cuentaperiodo += 1
    nombre_archivo = str(lper.id) + '.xlsx'
    directory = os.path.join(MEDIA_ROOT, 'auditoria', nombre_archivo)
    folder = os.path.join(os.path.join(SITE_STORAGE, 'media', 'auditoria', ''))
    rutaexcell = folder + str(nombre_archivo)
    if os.path.isfile(rutaexcell):
        os.remove(rutaexcell)
    # Inicializo cabecera de  excel
    __author__ = 'Unemi'
    wb = openxl.Workbook()
    ws = wb.active
    style_title = openxlFont(name='Arial', size=14, bold=True)
    style_cab = openxlFont(name='Arial', size=10, bold=True)
    alinear = alin(horizontal="center", vertical="center")
    response = HttpResponse(content_type="application/ms-excel")
    response['Content-Disposition'] = f'attachment; filename={lper.id}' + '-' + random.randint(1, 10000).__str__() + '.xlsx'
    ws.merge_cells('A1:E1')
    ws.merge_cells('A2:E2')
    ws['A1'] = 'UNIVERSIDAD ESTATAL DE MILAGRO'
    ws['A2'] = 'REPORTE DE AUDITORIA'
    celda1 = ws['A1']
    celda1.font = style_title
    celda1.alignment = alinear
    celda2 = ws['A2']
    celda2.font = style_title
    celda2.alignment = alinear
    columns = ['periodo',
               'carrera',
               'modalidad',
               'cedula',
               'genero',
               'alumno',
               'edad',
               'nivel_socio',
               'ppl',
               'provincia',
               'dis',
               'materia',
               'tipo_materia',
               'paralelo',
               'nivel',
               'exa1',
               'exa2',
               'nota_final',
               ]
    row_num = 3
    for col_num in range(0, len(columns)):
        celda = ws.cell(row=row_num, column=(col_num + 1), value=columns[col_num])
        celda.font = style_cab
    row_num = 4

    #Obtencion de directorio y varibles requeridas
    # directory = os.path.join(MEDIA_ROOT, 'auditoria', nombre_archivo)

    listadoalumnos = MateriaAsignada.objects.filter(materia__nivel__periodo_id=lper.id,
                                                    # matricula__inscripcion__persona__cedula='0942486861',
                                                    retiromanual=False,
                                                    matricula__retiradomatricula=False,
                                                    matricula__status=True,
                                                    materia__status=True,
                                                    materia__asignaturamalla__status=True,
                                                    materia__asignaturamalla__asignatura__modulo=False,
                                                    status=True).\
        exclude(materia__asignaturamalla__malla__carrera__coordinacion__id=9).\
        order_by('materia__nivel__periodo','materia__asignaturamalla__malla__carrera__nombre',
                 'materia__asignaturamalla__nivelmalla__id', 'matricula__inscripcion__persona__apellido1',
                 'matricula__inscripcion__persona__apellido2','materia__asignaturamalla__asignatura__nombre')

    # baseDate = datetime.today()
    totales = listadoalumnos.count()
    cuentamatri=0
    for lmatri in listadoalumnos:
        esppl = 'NO'
        nombreprovincia = ''
        if lmatri.matricula.inscripcion.persona.provincia:
            nombreprovincia = lmatri.matricula.inscripcion.persona.provincia.nombre
        if lmatri.matricula.inscripcion.persona.ppl:
            esppl = 'SI'
        nivelsocio = ''
        if lmatri.matricula.matriculagruposocioeconomico():
            nivelsocio=lmatri.matricula.matriculagruposocioeconomico().nombre

        disca = ''
        if lmatri.matricula.inscripcion.persona.tiene_discapasidad():
            disca=lmatri.matricula.inscripcion.persona.tiene_discapasidad()[0].tipodiscapacidad.nombre
        fechanace = ''
        if lmatri.matricula.inscripcion.persona.nacimiento:
            fechanacimiento = lmatri.matricula.inscripcion.persona.nacimiento
            if  fechanacimiento < hoy:
                edad = hoy.year - fechanacimiento.year - ((hoy.month, hoy.day) < (fechanacimiento.month, fechanacimiento.day))
                # print(edad)
        genero = ''
        if lmatri.matricula.inscripcion.persona.sexo:
            if lmatri.matricula.inscripcion.persona.sexo.id == 1:
                genero = 'MUJER'
            else:
                genero = 'HOMBRE'
        exa1=0
        exa2 = ''
        if lmatri.evaluaciongenerica_set.values("valor").filter(detallemodeloevaluativo__alternativa_id=20,status=True):
            listaexa1 = lmatri.evaluaciongenerica_set.filter(detallemodeloevaluativo__alternativa_id=20,status=True).order_by('detallemodeloevaluativo__nombre')[0]
            exa1=listaexa1.valor
            idexa1=listaexa1.id
            if lmatri.evaluaciongenerica_set.values("valor").filter(detallemodeloevaluativo__alternativa_id=20,status=True).exclude(pk=idexa1):
                listaexa2 = lmatri.evaluaciongenerica_set.filter(detallemodeloevaluativo__alternativa_id=20,status=True).exclude(pk=idexa1).order_by('-detallemodeloevaluativo__nombre')[0]
                exa2 = listaexa2.valor
        tipomate = ''
        if lmatri.materia.tipomateria:
            tipomate = lmatri.materia.get_tipomateria_display()
        ws.cell(row=row_num, column=1, value=lmatri.materia.nivel.periodo.__str__())
        ws.cell(row=row_num, column=2, value=lmatri.materia.asignaturamalla.malla.carrera.__str__())
        ws.cell(row=row_num, column=3, value=lmatri.materia.nivel.modalidad.__str__())
        ws.cell(row=row_num, column=4, value=lmatri.matricula.inscripcion.persona.cedula.__str__())
        ws.cell(row=row_num, column=5, value=genero)
        ws.cell(row=row_num, column=6, value=lmatri.matricula.inscripcion.persona.__str__())
        ws.cell(row=row_num, column=7, value=edad)
        ws.cell(row=row_num, column=8, value=nivelsocio.__str__())
        ws.cell(row=row_num, column=9, value=esppl.__str__())
        ws.cell(row=row_num, column=10, value=nombreprovincia.__str__())
        ws.cell(row=row_num, column=11, value=disca.__str__())
        ws.cell(row=row_num, column=12, value=lmatri.materia.asignaturamalla.asignatura.nombre)
        ws.cell(row=row_num, column=13, value=tipomate)
        ws.cell(row=row_num, column=14, value=lmatri.materia.paralelo)
        ws.cell(row=row_num, column=15, value=lmatri.materia.asignaturamalla.nivelmalla.__str__())
        ws.cell(row=row_num, column=16, value=exa1)
        ws.cell(row=row_num, column=17, value=exa2)
        ws.cell(row=row_num, column=18, value=lmatri.notafinal)
        row_num += 1
        cuentamatri += 1
        print('periodo: ' + str(cuentaperiodo) + ' de ' + str(totalperiodos) + ' periodos | registros: ' + str(cuentamatri) + ' de ' + str(totales))
    wb.save(directory)
