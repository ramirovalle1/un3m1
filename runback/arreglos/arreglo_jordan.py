import os
import sys
import io
import xlsxwriter
import xlwt
import openpyxl
import xlwt
from xlwt import *
from django.http import HttpResponse
from xlwt import *

# SITE_ROOT = os.path.dirname(os.path.realpath(__file__))
YOUR_PATH = os.path.dirname(os.path.realpath(__file__))
# print(f"YOUR_PATH: {YOUR_PATH}")
SITE_ROOT = os.path.dirname(os.path.dirname(YOUR_PATH))
SITE_ROOT = os.path.join(SITE_ROOT, '')
# print(f"SITE_ROOT: {SITE_ROOT}")
your_djangoproject_home = os.path.split(SITE_ROOT)[0]
# print(f"your_djangoproject_home: {your_djangoproject_home}")
sys.path.append(your_djangoproject_home)

from django.core.wsgi import get_wsgi_application

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "settings")
application = get_wsgi_application()

from django.http import HttpResponse
from settings import MEDIA_ROOT, BASE_DIR
from xlwt import easyxf, XFStyle
from sga.adm_criteriosactividadesdocente import asistencia_tutoria
from inno.models import *
from sga.models import *
from sagest.models import *
from balcon.models import *
from sga.funciones import convertirfecha, convertirfecha2
from Moodle_Funciones import crearhtmlphpmoodle
from sga.funciones import log, convertir_fecha, puede_realizar_accion, puede_realizar_accion_afirmativo, \
    null_to_decimal, generar_nombre, fechatope, convertir_fecha_invertida, variable_valor, MiPaginador, \
    dia_semana_ennumero_fecha, null_to_numeric, calculate_username, generar_usuario,generar_usuario_admision
from settings import MATRICULACION_LIBRE, UTILIZA_GRUPOS_ALUMNOS, NOMBRE_NIVEL_AUTOMATICO, MATRICULACION_POR_NIVEL, \
    CAPACIDAD_MATERIA_INICIAL, CUPO_POR_MATERIA, APROBACION_DISTRIBUTIVO, USA_EVALUACION_INTEGRAL, \
    TIPO_DOCENTE_TEORIA, TIPO_DOCENTE_PRACTICA, VERIFICAR_CONFLICTO_DOCENTE, TIPO_CUOTA_RUBRO, SITE_STORAGE, \
    HORAS_VIGENCIA, ADMISION_ID, USA_TIPOS_INSCRIPCIONES, NOTA_ESTADO_EN_CURSO, TIPO_INSCRIPCION_INICIAL, \
    TIPO_DOCENTE_FIRMA, TIPO_DOCENTE_AYUDANTIA, EMAIL_INSTITUCIONAL_AUTOMATICO, EMAIL_DOMAIN, ALUMNOS_GROUP_ID


def procesar_datos_titulo():
    with transaction.atomic():
        try:
            persona = Persona.objects.get(id=1)
            periodoactual = Periodo.objects.get(id=224)
            linea = 0
            miarchivo = openpyxl.load_workbook("CRUCE_TITULADOS.xlsx")



            # archivo_ = 'CRUCE_TITULADOS'
            # print("INICIO PROCESAMIENTO")
            # url_archivo = "{}/media/{}.xlsx".format(SITE_STORAGE, archivo_)
            # wb = openpyxl.load_workbook(filename=url_archivo)
            ws = miarchivo.get_sheet_by_name("CRUCE")
            worksheet = ws
            c = 0
            for row in worksheet.iter_rows(min_row=0):
                if linea >= 1:
                    currentValues, cadena = [], ''
                    for cell in row:
                        cadena += str(cell.value) + ' '
                        currentValues.append(str(cell.value))
                    identificacion = currentValues[0]



                    if Persona.objects.filter(cedula=identificacion, status=True).exists():
                        ePersona = Persona.objects.filter(cedula=identificacion, status=True)[0]


                        if Matricula.objects.filter(status=True, nivel__periodo__id=224, nivel__id__in=[1516,1517], cuposenescyt=True, inscripcion__persona=ePersona).exists():
                            matricula = Matricula.objects.get(status=True, nivel__periodo__id=224, nivel__id__in=[1516,1517], cuposenescyt=True, inscripcion__persona=ePersona)
                            eInscripcion = matricula.inscripcion

                            tipo_universidad = 0
                            tiene_titulo_publico = False
                            tipo_titulo_nacional=True
                            valor_x_materia = 20

                            if currentValues[6] == 'EXTRANJERO':
                                tipo_titulo_nacional = False


                            if currentValues[2] == 'Privada':
                                tipo_universidad = 4
                            if currentValues[2] == 'PÚBLICA':
                                tipo_universidad = 1
                            if currentValues[2] == 'PARTICULAR AUTOFINANCIADA':
                                tipo_universidad = 2
                            if currentValues[2] == 'PARTICULAR COFINANCIADA':
                                tipo_universidad = 3

                            nivel_estudio = currentValues[4]
                            tiponivel = 0
                            if nivel_estudio is None:
                                tiponivel = 0
                            if nivel_estudio in ['TECNICO_SUPERIOR', 'Técnico Superior', 'Tercer Nivel Técnico Superior', 'Tercer Nivel Tecnológico Superior Universitario']:
                                tiponivel = 2
                            if nivel_estudio in ['TERCER_NIVEL', 'Tercer Nivel o Pregrado', 'Tercer Nivel', 'Pregrado', 'Educación Superior de Grado o Tercer Nivel']:
                                tiponivel = 4
                            if nivel_estudio in ['CUARTO_NIVEL', 'Cuarto Nivel o Posgrado', 'Cuarto Nivel', 'Posgrado', 'Educación Superior de Posgrado o Cuarto Nivel']:
                                tiponivel = 5

                            if InstitucionEducacionSuperior.objects.filter(nombre__icontains=currentValues[1]).exists():
                                ies = InstitucionEducacionSuperior.objects.filter(nombre__icontains=currentValues[1])[0]
                            else:
                                ies = InstitucionEducacionSuperior(nombre=currentValues[1])
                                ies.save()

                            ePersonaTituloUniversidades = PersonaTituloUniversidad.objects.filter(
                                Q(nombrecarrera=currentValues[3]) | Q(codigoregistro=currentValues[8]), persona=ePersona)
                            if ePersonaTituloUniversidades.values("id").exists():
                                for ePersonaTituloUniversidad in ePersonaTituloUniversidades:
                                    ePersonaTituloUniversidad.persona = ePersona
                                    ePersonaTituloUniversidad.nombrecarrera = currentValues[3]
                                    ePersonaTituloUniversidad.codigoregistro = currentValues[8]
                                    ePersonaTituloUniversidad.tiponivel = tiponivel
                                    ePersonaTituloUniversidad.verificadosenescyt = True
                                    ePersonaTituloUniversidad.fechamigradosenescyt = datetime.now()

                                    ePersonaTituloUniversidad.universidad = ies
                                    ePersonaTituloUniversidad.tipouniversidad = tipo_universidad

                                    ePersonaTituloUniversidad.fecharegistro = convertirfecha2(currentValues[9])
                                    if not currentValues[10] == 'None':
                                        ePersonaTituloUniversidad.fecharegresado = convertirfecha2(currentValues[11])
                                    if not currentValues[11] == 'None':
                                        ePersonaTituloUniversidad.fechaacta = convertirfecha2(currentValues[10])
                                    if not currentValues[12] == 'None':
                                        ePersonaTituloUniversidad.fechainicio = convertirfecha2(currentValues[12])
                                    ePersonaTituloUniversidad.save(usuario_id=ePersona.usuario_id)

                            else:
                                ePersonaTituloUniversidad = PersonaTituloUniversidad(persona=ePersona)

                                ePersonaTituloUniversidad.nombrecarrera = currentValues[3]
                                ePersonaTituloUniversidad.codigoregistro = currentValues[8]
                                ePersonaTituloUniversidad.tiponivel = tiponivel
                                ePersonaTituloUniversidad.verificadosenescyt = True
                                ePersonaTituloUniversidad.fechamigradosenescyt = datetime.now()



                                ePersonaTituloUniversidad.universidad = ies
                                ePersonaTituloUniversidad.tipouniversidad = tipo_universidad
                                ePersonaTituloUniversidad.fecharegistro = convertirfecha2(currentValues[9])
                                if not currentValues[10] == 'None' and not currentValues[10] == 'NA':
                                    ePersonaTituloUniversidad.fecharegresado = convertirfecha2(currentValues[11])
                                if not currentValues[11] == 'None' and not currentValues[11] == 'NA':
                                    ePersonaTituloUniversidad.fechaacta = convertirfecha2(currentValues[10])
                                if not currentValues[12] == 'None' and not currentValues[12] == 'NA':
                                    ePersonaTituloUniversidad.fechainicio = convertirfecha2(currentValues[12])
                                ePersonaTituloUniversidad.save(usuario_id=ePersona.usuario_id)

                            if PersonaTituloUniversidad.objects.filter(status=True, persona=ePersona, tipouniversidad=1, tiponivel__in=[4,2]).exists() and tipo_titulo_nacional:
                                personaverificar=PersonaTituloUniversidad.objects.filter(status=True, persona=ePersona, tipouniversidad=1, tiponivel__in=[4,2]).last()
                                if personaverificar.fechainicio.year > 2008 and personaverificar.tiponivel == 4:
                                    tiene_titulo_publico = True
                                elif personaverificar.fecharegistro.year > 2019 and personaverificar.tiponivel == 2:
                                    tiene_titulo_publico = True





                            #consultar títulos tecnicos y tecnologicos y financiamiento
                            if tiene_titulo_publico:


                                if PerdidaGratuidad.objects.filter(inscripcion=eInscripcion).exists():
                                    ePerdidaGratuidad = PerdidaGratuidad.objects.get(inscripcion=eInscripcion)
                                    ePerdidaGratuidad.titulo_sniese = tiene_titulo_publico
                                    ePerdidaGratuidad.observacion = f"{ePerdidaGratuidad.observacion}, Registra TITULO TERCER NIVEL REGISTRO SNIESE"
                                    ePerdidaGratuidad.save(usuario_id=persona.usuario.id)
                                    eInscripcion.gratuidad = False
                                    eInscripcion.estado_gratuidad = 3
                                    eInscripcion.save(usuario_id=persona.usuario.id)

                                else:
                                    observacion = f"Reportado por la SENESCYT"
                                    observacion = f"{observacion}, Registra TITULO TERCER NIVEL REGISTRO SNIESE"
                                    ePerdidaGratuidad = PerdidaGratuidad(inscripcion=eInscripcion)
                                    ePerdidaGratuidad.motivo = 1
                                    ePerdidaGratuidad.titulo = None
                                    ePerdidaGratuidad.titulo_sniese = tiene_titulo_publico
                                    ePerdidaGratuidad.segunda_carrera_raes = False
                                    ePerdidaGratuidad.observacion = observacion
                                    ePerdidaGratuidad.save(usuario_id=persona.usuario.id)
                                    eInscripcion.gratuidad = False
                                    eInscripcion.estado_gratuidad = 3
                                    eInscripcion.save(usuario_id=persona.usuario.id)



                            if eInscripcion.estado_gratuidad == 3:
                                num_materias = MateriaAsignada.objects.filter(status=True, matricula=matricula).count()
                                if eInscripcion.sesion_id == 13:
                                    tiporubromatricula = TipoOtroRubro.objects.get(pk=3019)
                                else:
                                    tiporubromatricula = TipoOtroRubro.objects.get(pk=3011)

                                if num_materias > 0:
                                    valor_total = num_materias * valor_x_materia
                                    matricula.estado_matricula = 1
                                    matricula.save(usuario_id=persona.usuario.id)
                                    rubro1 = Rubro.objects.filter(persona=eInscripcion.persona,
                                                                  matricula=matricula, status=True).first()

                                    if not rubro1:
                                        print(f"Creando Rurbo - {ePersona.cedula}")
                                        rubro1 = Rubro(tipo=tiporubromatricula,
                                                       persona=eInscripcion.persona,
                                                       matricula=matricula,
                                                       nombre=tiporubromatricula.nombre + ' - ' + periodoactual.nombre,
                                                       cuota=1,
                                                       fecha=datetime.now().date(),
                                                       # fechavence=datetime.now().date() + timedelta(days=22),
                                                       fechavence=datetime.now().date(),
                                                       valor=valor_total,
                                                       iva_id=1,
                                                       valoriva=0,
                                                       valortotal=valor_total,
                                                       saldo=valor_total,
                                                       cancelado=False)
                                        rubro1.save(usuario_id=persona.usuario.id)

                            c += 1
                            print(f"{c} actualizado")
                linea += 1
            print(f"FIN - {c} registros actualizados")

        except Exception as ex:
            transaction.set_rollback(True)
            textoerror = '{} Linea:{} '.format(str(ex), sys.exc_info()[-1].tb_lineno)
            print(textoerror)

procesar_datos_titulo()

