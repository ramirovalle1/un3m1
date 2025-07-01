import io
import os
import sys
import xlsxwriter
import xlwt
import openpyxl
from xlwt import *

# SITE_ROOT = os.path.dirname(os.path.realpath(__file__))
YOUR_PATH = os.path.dirname(os.path.realpath(__file__))
# print(f"YOUR_PATH: {YOUR_PATH}")
SITE_ROOT = os.path.dirname(os.path.dirname(YOUR_PATH))
SITE_ROOT = os.path.join(SITE_ROOT, '')
# print(f"SITE_ROOT: {SITE_ROOT}")
your_djangoproject_home = os.path.split(SITE_ROOT)[0]
# print(f"your_djangoproject_home: {your_djangoproject_home}")
sys.path.append(your_djangoproject_home)

from django.core.wsgi import get_wsgi_application

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "settings")
application = get_wsgi_application()

from django.http import HttpResponse
from settings import MEDIA_ROOT, BASE_DIR
from xlwt import easyxf, XFStyle
from sga.adm_criteriosactividadesdocente import asistencia_tutoria
from inno.models import *
from sga.models import *
from sagest.models import *
from inno.funciones import *
import concurrent.futures
from Moodle_Funciones import CrearExamenMoodle
#
# try:
#     for malla in Malla.objects.filter(status=True, carrera__coordinacion__id__in=[1, 2, 3, 4, 5]).distinct():
#         for asignaturamalla in AsignaturaMalla.objects.filter(status=True, malla=malla):
#             if not DetalleAsignaturaMallaModalidad.objects.filter(status=True, asignaturamalla=asignaturamalla).exists():
#                 detasigmdod= DetalleAsignaturaMallaModalidad(
#                     asignaturamalla= asignaturamalla,
#                     modalidad= 1)
#
#                 detasigmdod.save()
#
# except Exception as ex:
#     print(ex)
#
# print("fin")


#importar nota ingles
# try:
#
#     alumno = MateriaAsignada.objects.get(pk=1897203)
#     matricula = alumno.matricula
#     if not alumno.materia.notas_de_moodle(alumno.matricula.inscripcion.persona):
#         raise NameError(u"Lo sentimos, no se encuentra notas del examen.")
#     for notasmooc in alumno.materia.notas_de_moodle(alumno.matricula.inscripcion.persona):
#         campo = alumno.campo(notasmooc[1].upper())
#         if Decimal(notasmooc[0]) <= 0:
#             raise NameError(u"Lo sentimos, no se puede importar notas con cero.")
#         if type(notasmooc[0]) is Decimal:
#             if null_to_decimal(campo.valor) != float(notasmooc[0]) or (alumno.asistenciafinal < campo.detallemodeloevaluativo.modelo.asistenciaaprobar):
#                 actualizar_nota_planificacion(alumno.id, notasmooc[1].upper(), notasmooc[0])
#                 auditorianotas = AuditoriaNotas(evaluaciongenerica=campo, manual=False, calificacion=notasmooc[0])
#                 auditorianotas.save()
#         else:
#             if null_to_decimal(campo.valor) != float(0) or (alumno.asistenciafinal < campo.detallemodeloevaluativo.modelo.asistenciaaprobar):
#                 actualizar_nota_planificacion(alumno.id, notasmooc[1].upper(), notasmooc[0])
#                 auditorianotas = AuditoriaNotas(evaluaciongenerica=campo, manual=False, calificacion=0)
#                 auditorianotas.save()
#     alumno.importa_nota = True
#     alumno.cerrado = True
#     alumno.fechacierre = datetime.now().date()
#     alumno.save(actualiza=False)
#     d = locals()
#     exec(alumno.materia.modeloevaluativo.logicamodelo, globals(), d)
#     d['calculo_modelo_evaluativo'](alumno)
#     alumno.cierre_materia_asignada()
#
# except Exception as ex:
#     print(ex)

# def reajustar_asistencia_dia_especifico():
#     # fechas = [
#     #     (datetime(2021, 11, 29, 0, 0, 0)).date(),
#     #     (datetime(2021, 11, 30, 0, 0, 0)).date(),
#     #     (datetime(2021, 12, 1, 0, 0, 0)).date(),
#     #     (datetime(2021, 12, 2, 0, 0, 0)).date()
#     # ]
#     fechas = [
#         # (datetime(2021, 11, 29, 0, 0, 0)).date(),
#         # (datetime(2021, 11, 30, 0, 0, 0)).date(),
#         # (datetime(2021, 12, 1, 0, 0, 0)).date(),
#         # (datetime(2021, 12, 2, 0, 0, 0)).date(),
#         # (datetime(2021, 12, 3, 0, 0, 0)).date(),
#         # (datetime(2021, 12, 4, 0, 0, 0)).date(),
#         # (datetime(2021, 12, 5, 0, 0, 0)).date()
#         # (datetime(2021, 12, 16, 0, 0, 0)).date(),
#         # (datetime(2021, 12, 17, 0, 0, 0)).date()
#         (datetime(2022, 4, 1, 0, 0, 0)).date(),
#         (datetime(2022, 4, 8, 0, 0, 0)).date()
#     ]
#     ePeriodo = Periodo.objects.get(pk=119)
#     ePeriodoAcademia = ePeriodo.get_periodoacademia()
#     # fecha = (datetime(2021, 11, 29, 0, 0, 0)).date()
#     cedulas = []
#
#     for fecha in fechas:
#         ids_p = []
#         dia = (fecha.weekday() + 1)
#         # for doc in cedulas:
#         #     if Persona.objects.values("id").filter(Q(cedula=doc) | Q(pasaporte=doc)).exists():
#         #         ids_p.append(Persona.objects.filter(Q(cedula=doc) | Q(pasaporte=doc))[0].id)
#         # profesores = Profesor.objects.filter(categoria_id__in=[4, 5, 6], profesormateria__materia__nivel__periodo=ePeriodo, profesormateria__activo=True).distinct()
#         ids_persona = [145717]
#         # for cedula in cedulas:
#         #     if Persona.objects.filter(Q(cedula=cedula) | Q(pasaporte=cedula)).exists():
#         #         ePersona = Persona.objects.filter(Q(cedula=cedula) | Q(pasaporte=cedula))[0]
#         #         ids_persona.append(ePersona.id)
#         profesores = Profesor.objects.filter(profesormateria__materia__nivel__periodo=ePeriodo, profesormateria__activo=True, persona_id__in=ids_persona).distinct()
#         # profesores = Profesor.objects.filter(profesormateria__materia__nivel__periodo=ePeriodo, profesormateria__activo=True).distinct()
#         # profesores = Profesor.objects.filter(pk=903).distinct()
#         total_profesores = profesores.count()
#         print("*** FECHA A PROCESAR: " + fecha.__str__() + "\r")
#         contP = 1
#         for profesor in profesores:
#             print(f"***** ({contP}/{total_profesores}) -> Profesor: {profesor.__str__()}")
#             # clases = Clase.objects.filter(activo=True, inicio__lte=fecha, fin__gte=fecha, dia=dia, status=True, materia__nivel__periodo_id=119, profesor_id=profesor.id).exclude()
#             clases = Clase.objects.filter(activo=True, inicio__lte=fecha, fin__gte=fecha, dia=dia, status=True,
#                                           materia__nivel__periodo=ePeriodo,
#                                           materia__nivel__periodo__visible=True,
#                                           materia__nivel__periodo__visiblehorario=True,
#                                           materia__fechafinasistencias__gte=fecha,
#                                           materia__profesormateria__profesor_id=profesor.id,
#                                           profesor_id=profesor.id,
#                                           materia__profesormateria__tipoprofesor_id__in=[10, 11, 12, 1, 5, 8, 7, 2, 14, 15],
#                                           tipoprofesor_id__in=[10, 11, 12, 1, 5, 8, 7, 2, 14, 15])
#             clases = clases.filter(Q(turno__comienza__gte=time(7, 0, 0), turno__termina__lte=time(10, 59, 00))|Q(turno__comienza__gte=time(11, 0, 0), turno__termina__lte=time(15, 30, 00))|Q(turno__comienza__gte=time(10, 45, 0), turno__termina__lte=time(11, 44, 00))|Q(turno__comienza__gte=time(11, 45, 0), turno__termina__lte=time(12, 44, 00))|Q(turno__comienza__gte=time(13, 00, 0), turno__termina__lte=time(13, 59, 00))|Q(turno__comienza__gte=time(14, 00, 0), turno__termina__lte=time(14, 59, 00))|Q(turno__comienza__gte=time(15, 30, 0), turno__termina__lte=time(16, 29, 00)) )
#             total_clases = clases.count()
#             contC = 1
#             for cl in clases:
#                 print(f"******** ({contC}/{total_clases}) -> Clase: {cl.__str__()}")
#                 # if cl.materia.profesor_principal():
#                 if LeccionGrupo.objects.values("id").filter(profesor=profesor, turno=cl.turno, fecha=fecha).exists():
#                     lecciongrupo = LeccionGrupo.objects.get(profesor=profesor, turno=cl.turno, fecha=fecha)
#                     lecciongrupo.abierta = False
#                     lecciongrupo.save()
#                 else:
#                     lecciongrupo = LeccionGrupo(profesor=profesor,
#                                                 turno=cl.turno,
#                                                 aula=cl.aula,
#                                                 dia=cl.dia,
#                                                 fecha=fecha,
#                                                 horaentrada=cl.turno.comienza,
#                                                 horasalida=cl.turno.termina,
#                                                 abierta=False,
#                                                 automatica=True,
#                                                 contenido='REGISTRO MASIVO 2022  - AUTORIZADO POR DIRECTOR TICS',
#                                                 observaciones='REGISTRO MASIVO 2022 - AUTORIZADO POR DIRECTOR TICS')
#                     lecciongrupo.save()
#                 if Leccion.objects.values("id").filter(clase=cl, fecha=fecha).exists():
#                     leccion = Leccion.objects.get(clase=cl, fecha=fecha)
#                     leccion.abierta = False
#                     leccion.save()
#                 else:
#                     leccion = Leccion(clase=cl,
#                                       fecha=fecha,
#                                       horaentrada=cl.turno.comienza,
#                                       horasalida=cl.turno.termina,
#                                       abierta=False,
#                                       contenido=lecciongrupo.contenido,
#                                       observaciones=lecciongrupo.observaciones)
#                     leccion.save()
#                 if not lecciongrupo.lecciones.values("id").filter(pk=leccion.id).exists():
#                     lecciongrupo.lecciones.add(leccion)
#
#                 asignados = None
#                 # SE FILTRA SI LA MATERIA TIENE TIPO DE PROFESOR PRACTICA Y LA CLASE TAMBIEN
#                 # 1 => CLASE PRESENCIAL
#                 # 2 => CLASE VIRTUAL SINCRÓNICA
#                 # 8 => CLASE REFUERZO SINCRÓNICA
#                 if cl.tipoprofesor.id == 2 and cl.tipohorario in [1, 2, 8]:
#                     if cl.grupoprofesor:
#                         if cl.grupoprofesor.paralelopractica:
#                             # grupoprofesor_id = clase.grupoprofesor.id
#                             if cl.grupoprofesor.listado_inscritos_grupos_practicas().exists():
#                                 listado_alumnos_practica = cl.grupoprofesor.listado_inscritos_grupos_practicas()
#                                 if ePeriodoAcademia.valida_asistencia_pago:
#                                     asignados = MateriaAsignada.objects.filter(pk__in=listado_alumnos_practica.values_list('materiaasignada_id', flat=True), matricula__estado_matricula__in=[2, 3]).distinct()
#                                 else:
#                                     asignados = MateriaAsignada.objects.filter(pk__in=listado_alumnos_practica.values_list('materiaasignada_id', flat=True)).distinct()
#                             else:
#                                 raise print(u"Clase no creada, de tipo práctica no tiene alumnos. %s" % cl)
#
#                         else:
#                             raise print(u"Clase no creada, de tipo práctica no tiene paralelos. %s" % cl)
#                     else:
#                         raise NameError(u"Clase no creada, de tipo práctica no tiene grupos. %s" % cl)
#                 else:
#                     asignados = cl.materia.asignados_a_esta_materia()
#
#                 total_asistencias = 0
#                 if AsistenciaLeccion.objects.values("id").filter(leccion=leccion).exists():
#                     total_asistencias = AsistenciaLeccion.objects.filter(leccion=leccion).count()
#                     for asis in AsistenciaLeccion.objects.filter(leccion=leccion):
#                         # if not asis.asistio:
#                         asis.asistio = True
#                         asis.save()
#                         mateasig = asis.materiaasignada
#                         mateasig.save(actualiza=True)
#                         mateasig.actualiza_estado()
#                 else:
#                     total_asistencias = asignados.count()
#                     for materiaasignada in asignados:
#                         if not AsistenciaLeccion.objects.values("id").filter(leccion=leccion, materiaasignada=materiaasignada).exists():
#                             asistencialeccion = AsistenciaLeccion(leccion=leccion,
#                                                                   materiaasignada=materiaasignada,
#                                                                   asistio=True)
#                             asistencialeccion.save()
#                             materiaasignada.save(actualiza=True)
#                             materiaasignada.actualiza_estado()
#                         # guardar temas de silabo
#                 lecciongrupo.save()
#                 print(f"*********** (Total Asistencia: {total_asistencias})")
#                 # print(cl)
#                 contC += 1
#
#             contP += 1
#
#
# reajustar_asistencia_dia_especifico()


##mostrar horario solo en la coordinacion de salud

#
# print("fin")

#

# try:
#     materias = Materia.objects.filter(status=True, nivel__periodo=126)
#     for materia in materias:
#         if materia.coordinacion().id != 1:
#             materia.visiblehorario = False
#             materia.save()
#             print(materia.id)
# except Exception as ex:
#     print(ex)
#
#
# try:
#     mallas = Malla.objects.filter(status=True, carrera__coordinacion__id__in=[1, 2, 3, 4, 5]).exclude(carrera__id__in=[157, 90, 129])
#     for malla in mallas:
#         if malla.inicio.year == 2012 :
#             print(str(malla.id) + ",")
# except Exception as ex:
#     print(ex)

#
#
# def reasjute_horas_creditos_malla_modulos_ingles():
#
#     print(SITE_STORAGE)
#     with xlsxwriter.Workbook(f'{SITE_STORAGE}/media/migracion_modulos_ingles_2021_{random.randrange(1, 100)}.xlsx') as workbook:
#         wk = workbook.add_worksheet(f"Hoja1")
#         columns = [
#             (u"ID_INSCRIPCION", 10),
#             (u"APELLIDOS_NOMBRES", 100),
#             (u"CEDULA", 80),
#             (u"CARRERA", 80),
#             (u"ID_MALLA", 80),
#             (u"ID_RECORD", 10),
#             (u"MODULO", 30),
#             (u"CREDITOS_ANTERIOR", 10),
#             (u"CREDITOS_ACTUAL", 10),
#             (u"HORAS_ANTERIOR", 10),
#             (u"HORAS_ACTUAL", 10),
#         ]
#         row_num = 1
#         for col_num in range(len(columns)):
#             wk.write(row_num, col_num, columns[col_num][0])
#             wk.set_column(col_num, col_num, columns[col_num][1])
#         row_num = 2
#         wrap_format = workbook.add_format({'text_wrap': True})
#         c = 0
#         ids = [10, 21, 15, 147, 16, 148, 12, 232, 228, 233, 18, 17, 222, 225, 202, 226, 205, 8, 207, 6, 210, 201, 224, 356, 9, 204, 218, 206, 212, 114, 219, 113, 4, 5, 3, 7, 115, 199, 237, 208, 332, 231, 200, 11, 19, 14, 213, 258, 355, 173, 174, 13, 2, 172]
#        # ids = [210]
#       #2013 -2012  ids = [9, 10, 13, 2, 15, 19, 11, 7, 3, 14, 16, 115, 6, 12, 148, 147, 8, 4, 113, 18, 17, 114, 5, 21]
#         # inscripcion_ids = [131795, 104493]
#         # ids = [219]
#         mallas = Malla.objects.filter(pk__in=ids)
#         for malla in mallas:
#             print(f"Malla: {malla.__str__()}")
#             for inscripcionmalla in InscripcionMalla.objects.filter(malla=malla):
#                 inscripcion = inscripcionmalla.inscripcion
#                 if not inscripcion.usado_graduados() or not inscripcion.egresado():
#                     # print(f"Inscripcion: {inscripcion.__str__()}")
#                     recordacademico = inscripcion.recordacademico_set.filter(modulomalla__isnull=False, aprobada=True, creditos=0)
#                     # print(f"Total de record: {len(recordacademico)}")
#                     for record in recordacademico:
#                         print(f"Inscripcion: {inscripcion.__str__()}")
#                         record.actualizar()
#                         historico = record.mi_historico()
#                         historico.creditos = record.modulomalla.creditos
#                         historico.horas = record.modulomalla.horas
#                         historico.validapromedio = False
#                         historico.valida = True
#                         historico.save()
#                         creditos_anterior = record.creditos
#                         creditos_actual = record.modulomalla.creditos
#                         record.creditos = record.modulomalla.creditos
#                         horas_anterior = record.horas
#                         horas_actual = record.modulomalla.horas
#                         record.horas = record.modulomalla.horas
#                         record.validapromedio = False
#                         record.valida = True
#                         record.save()
#                         # ID_INSCRIPCION
#                         wk.write(row_num, 0, inscripcion.id)
#                         # APELLIDOS_NOMBRES
#                         wk.write(row_num, 1, inscripcion.persona.nombre_completo_inverso())
#                         # CEDULA
#                         wk.write(row_num, 2, inscripcion.persona.documento())
#                         # CARRERA
#                         wk.write(row_num, 3, inscripcion.carrera.__str__())
#                         # ID_MALLA
#                         wk.write(row_num, 4, malla.id)
#                         # ID_RECORD
#                         wk.write(row_num, 5, record.id)
#                         # MODULO
#                         wk.write(row_num, 6, record.modulomalla.asignatura.nombre)
#                         # CREDITOS_ANTERIOR
#                         wk.write(row_num, 7, creditos_anterior)
#                         # CREDITOS_ACTUAL
#                         wk.write(row_num, 8, creditos_actual)
#                         # HORAS_ANTERIOR
#                         wk.write(row_num, 9, horas_anterior)
#                         # HORAS_ACTUAL
#                         wk.write(row_num, 10, horas_actual)
#                         row_num += 1
#                     print(f"Final de Inscripcion: {inscripcion.__str__()}")
#                     print(f"------------------------------------------------")
#
# #
# reasjute_horas_creditos_malla_modulos_ingles()
# #
#
# try:
#     linea = 1
#     for inscripcion in Inscripcion.objects.filter(status=True):
#         # print('%s - %s-%s' % (linea, inscripcion.id, inscripcion))
#         inscripcionold = inscripcion.inscripcionold
#         asigmal = AsignaturaMalla.objects.get(asignatura_id=1053, malla_id=32)
#         # for asigmal in AsignaturaMalla.objects.filter(malla_id=32):
#         modulos = RecordAcademico.objects.filter(Q(aprobada=True) | Q(noaplica=True), status=True, inscripcion=inscripcion, asignatura_id=asigmal.asignatura_id, creditos=0).order_by('id')
#         for record2 in modulos:
#                 print('%s - %s-%s' % (linea, inscripcion.id, inscripcion))
#                 record2.creditos = asigmal.creditos
#                 record2.save()
#                 historico =  HistoricoRecordAcademico.objects.filter(status=True, recordacademico = record2).order_by('-aprobada', '-fecha')[0]
#                 historico.creditos = record2.creditos
#                 historico.save()
#                 # record2.actualizar()
#                 print('------------------ COMPUP %s - %s' % (linea, record2.asignatura))
#
#         linea += 1
# except Exception as ex:
#     print('error: %s' % ex)

#
#
#
#
#
# try:
#     linea = 1
#     for inscripcion in Inscripcion.objects.filter(pk__in=[57183, 29627]):
#         # print('%s - %s-%s' % (linea, inscripcion.id, inscripcion))
#         inscripcion = inscripcion
#         asigmal = AsignaturaMalla.objects.get(asignatura_id=1053, malla_id=32)
#         # for asigmal in AsignaturaMalla.objects.filter(malla_id=32):
#         modulos = RecordAcademico.objects.filter(Q(aprobada=True) | Q(noaplica=True), status=True, inscripcion=inscripcion, asignatura_id=asigmal.asignatura_id, creditos=0).order_by('id')
#         for record2 in modulos:
#                 print('%s - %s-%s' % (linea, inscripcion.id, inscripcion))
#                 record2.creditos = asigmal.creditos
#                 record2.save()
#                 historico =  HistoricoRecordAcademico.objects.filter(status=True, recordacademico = record2).order_by('-aprobada', '-fecha')[0]
#                 historico.creditos = record2.creditos
#                 historico.save()
#                 # record2.actualizar()
#                 print('------------------ COMPUP %s - %s' % (linea, record2.asignatura))
#
#         linea += 1
# except Exception as ex:
#     print('error: %s' % ex)
#
#
#
#







#matricula curricular
# try:
#     ids = [440277, 455571, 478125, 463361, 464563, 455571, 464631, 462920, 463037, 473409, 454323, 469881, 470060, 470302, 470455, 470457, 470611, 436241, 474759,
#            468528, 475939]
#     matriculas = Matricula.objects.filter(pk__in=ids)
#     for matricula in matriculas:
#         materiaasignada = MateriaAsignada.objects.filter(status=True, matricula=matricula,materia__asignaturamalla__validarequisitograduacion=True, materia__asignaturamalla__nivelmalla_id=8)
#         for matasig in materiaasignada:
#             matasig.status=False
#             matasig.save()
#             print(matasig.id, matricula.id)
#             asunto = u"RETIRO DE MATERIA DE INTEGRACIÓN CURRICULAR"
#             # correo = []
#             correo = matricula.inscripcion.persona.lista_emails_2()
#             correo.append('mleong2@unemi.edu.ec')
#             if matricula.inscripcion.persona.emailinst:
#                 send_html_mail(asunto, "emails/notificar_retiro_materia.html",
#                                {
#                                    'sistema': u'SGA - UNEMI',
#                                    'estudiante': matricula.inscripcion.persona
#                                },
#                                correo, [],
#                                cuenta=CUENTAS_CORREOS[0][1])
#                 print("Correo enviado %s +++++ %s " % (matricula.inscripcion.persona, matricula.inscripcion.persona.emailinst))
#                 # time.sleep(2)
#             else:
#                 print("No tiene correo %s +++++  " % (matricula.inscripcion.persona))
#
#             # matasig.delete()
#
#
# except Exception as ex:
#     print('error: %s' % ex)

#cambio tipo de profesor en horario
# try:
#     profesormateria = ProfesorMateria.objects.filter(materia__nivel__periodo=126, tipoprofesor_id=13, status=True)
#     for pm in profesormateria:
#         horarios = Clase.objects.filter(status=True, profesor=pm.profesor, materia=pm.materia, materia__nivel__periodo=126, tipoprofesor=2)
#         for horario in horarios:
#             horario.tipoprofesor_id = pm.tipoprofesor_id
#             horario.save()
#             print("Cambio tipo profesor en horario  %s - %s - %s +++++  " % (horario.id, horario.profesor.persona, horario.materia ))
#
#
#     horarios2 = Clase.objects.filter(status=True, tipoprofesor_id=13, materia__nivel__periodo=126).exclude(aula_id = 265)
#     for hor in horarios2:
#         if hor.aula_id != 265 :
#             hor.aula_id = 265
#             hor.save()
#
#
# except Exception as ex:
#     print('error: %s' % ex)
#
# print("fin")


#cambio aula horario
# try:
#     horarios = Clase.objects.filter(status=True, tipoprofesor_id=13, materia__nivel__periodo=126)
#     for horario in horarios:
#         if horario.aula_id != 265 :
#             horario.aula_id = 265
#             horario.save()
#
# except Exception as ex:
#     print('error: %s' % ex)
#
# print("fin")

#actualizar horas
# try:
#     profdistri = ProfesorDistributivoHoras.objects.filter(periodo=126, status=True, coordinacion__id__in=[1, 2, 3 , 4, 5])
#     for pd in profdistri:
#             profemat = ProfesorMateria.objects.filter(profesor = pd.profesor, materia__nivel__periodo=126)
#             if profemat.exists():
#                 horasclase = ProfesorMateria.objects.filter(profesor=pd.profesor, materia__nivel__periodo=126, activo=True,
#                                               principal=True, status=True).aggregate(horas=Sum('hora')).get('horas')
#                 if horasclase:
#                     detalle = DetalleDistributivo.objects.filter(status=True, distributivo = pd, criteriodocenciaperiodo = 675)
#                     if detalle.exists():
#                         detalle = DetalleDistributivo.objects.get(status=True, distributivo = pd, criteriodocenciaperiodo = 675)
#                         if detalle.horas != horasclase:
#                             print("Actualización horas profesor %s - %s por: %s  +++++  " % (str(detalle.distributivo.profesor), detalle.horas, horasclase ))
#                             detalle.horas = horasclase
#                             detalle.save()
#
#
# except Exception as ex:
#     print('error: %s' % ex)


# try:
#     profesor = ProfesorMateria.objects.filter(status =True, materia__nivel__periodo = 119,tipoprofesor_id= 8)
#
#     for prof in profesor:
#         print(prof.profesor)
#
# except Exception as ex:
#     print('error: %s' % ex)
#
# print("fin")

#cambio de estado de compendio de estado "ingresado" a "revisión crai"
# try:
#     compendios = CompendioSilaboSemanal.objects.filter(status=True, silabosemanal__silabo__materia__nivel__periodo = 126, estado_id=1)
#     for compendio in compendios:
#             nombre = str(compendio.archivocompendio)
#             ext = nombre[nombre.rfind("."):]
#             if ext == '.doc' or ext == '.docx':
#                 compendio.estado_id = 5
#                 compendio.save()
# except Exception as ex:
#     print('error: %s' % ex)
#
# print("fin")
def liquidar_deuda():
    try:
        bloqueados = Matricula.objects.filter(status=True,nivel__periodo_id=153, inscripcion__persona__cedula__in=['0918710674'])
        idpersona_liquidar_deuda= []
        for bloqueado in bloqueados:
            rubros = Rubro.objects.filter(status=True, matricula_id=bloqueado.id)
            totalrubros = Rubro.objects.filter(status=True, matricula_id=bloqueado.id).count()
            totalcancelado = 0
            for rubro in rubros:
                if rubro.cancelado:
                    totalcancelado += 1
            if totalcancelado == totalrubros:
                idpersona_liquidar_deuda.append(bloqueado.inscripcion.persona.id)


        matriculados_periodo_actual = Matricula.objects.filter(status=True, nivel__periodo_id=177, inscripcion__persona__id__in = idpersona_liquidar_deuda)
        for matriculadoactual in matriculados_periodo_actual:

            rubros_periodo_actual = Rubro.objects.filter(status=True, matricula_id=matriculadoactual.id, cancelado=False)
            if rubros_periodo_actual:
                print('liquidar rubro -', matriculadoactual.inscripcion.persona.cedula)
                for rubroact in rubros_periodo_actual:
                    if not rubroact.cantidad_pagos() > 0:
                        if not rubroact.bloqueado:
                            subtotal0 = 0
                            subtotaliva = 0
                            iva = 0
                            if rubroact.iva.porcientoiva > 0:
                                subtotaliva = Decimal(rubroact.saldo / (rubroact.iva.porcientoiva + 1)).quantize(
                                    Decimal('.01'))
                                iva = Decimal(rubroact.saldo - subtotaliva).quantize(Decimal('.01'))
                            else:
                                subtotal0 = rubroact.saldo

                            pago = Pago(rubro=rubroact,
                                        fecha=datetime.now().date(),
                                        subtotal0=subtotal0,
                                        subtotaliva=subtotaliva,
                                        iva=iva,
                                        valordescuento=0,
                                        valortotal=rubroact.saldo,
                                        efectivo=False)
                            pago.save()
                            liquidacion = PagoLiquidacion(fecha=datetime.now().date(),
                                                          motivo='Liquidación de rubro según resolución OCS-SO-6-2023-No8',
                                                          valor=rubroact.saldo)
                            liquidacion.save()
                            liquidacion.pagos.add(pago)
                            rubroact.save()
                            matriculadoactual.estado_matricula = 2
                            matriculadoactual.save()
                        else:
                            print('tiene rubro bloqueado',matriculadoactual.inscripcion.persona.cedula)
                    else:
                        print('tiene rubro con pagos', matriculadoactual.inscripcion.persona.cedula )


    except Exception as ex:
        print(ex)
        transaction.set_rollback(True)


def liquidar_deuda_movilidad():
    try:
        rubros_periodo_actual = Rubro.objects.filter(status=True, matricula_id=670834, cancelado=False)
        matriculado = Matricula.objects.get(status=True, pk=670834)
        if rubros_periodo_actual:
            print('liquidar rubro - 734710364')
            for rubroact in rubros_periodo_actual:
                if not rubroact.cantidad_pagos() > 0:
                    if not rubroact.bloqueado:
                        subtotal0 = 0
                        subtotaliva = 0
                        iva = 0
                        if rubroact.iva.porcientoiva > 0:
                            subtotaliva = Decimal(rubroact.saldo / (rubroact.iva.porcientoiva + 1)).quantize(
                                    Decimal('.01'))
                            iva = Decimal(rubroact.saldo - subtotaliva).quantize(Decimal('.01'))
                        else:
                            subtotal0 = rubroact.saldo

                        pago = Pago(rubro=rubroact,
                                        fecha=datetime.now().date(),
                                        subtotal0=subtotal0,
                                        subtotaliva=subtotaliva,
                                        iva=iva,
                                        valordescuento=0,
                                        valortotal=rubroact.saldo,
                                        efectivo=False)
                        pago.save()
                        liquidacion = PagoLiquidacion(fecha=datetime.now().date(),
                                                          motivo='Liquidación de rubro por movilidad internacional',
                                                          valor=rubroact.saldo)
                        liquidacion.save()
                        liquidacion.pagos.add(pago)
                        rubroact.save()
                        matriculado.estado_matricula = 2
                        matriculado.save()
                    else:
                        print('tiene rubro bloqueado')
                else:
                    print('tiene rubro con pagos')


    except Exception as ex:
        print(ex)
        transaction.set_rollback(True)



def calificacion_transversales_domingo_presenciales_1():
    with transaction.atomic():
        try:
            periodo = Periodo.objects.get(id=177)
            asignaturas = DetalleGrupoAsignatura.objects.values_list('asignatura_id', flat=True).filter(status=True, grupo_id=1)
            materiasinvestigacion = MateriaAsignada.objects.filter(status=True,
                                                              materia__asignatura_id__in=asignaturas,
                                                              matricula__nivel__periodo=periodo,
                                                              matricula__bloqueomatricula=False,
                                                              matricula__retiradomatricula=False, materia__status=True,
                                                              matricula__status=True,
                                                              matricula__inscripcion__carrera__modalidad__in=[1],
                                                              matricula__inscripcion__carrera_id__in=[140],
                                                              materia__modeloevaluativo_id=27)
            cursor = connections['db_moodle_virtual'].cursor()
            for materiaasignada in materiasinvestigacion:
                # guardo_nota = False
                usuario = materiaasignada.matricula.inscripcion.persona.usuario.username
                #SE NECESITA EL ID DE CURSO MOODLE
                sql = """
                                    SELECT ROUND(nota.finalgrade,2), UPPER(gc.fullname)
                                            FROM mooc_grade_grades nota
                                    INNER JOIN mooc_grade_items it ON nota.itemid=it.id AND courseid=2039 AND itemtype='category'
                                    INNER JOIN mooc_grade_categories gc ON gc.courseid=it.courseid AND gc.id=it.iteminstance AND gc.depth=2
                                    INNER JOIN mooc_user us ON nota.userid=us.id
                                    WHERE us.username = '%s' and not UPPER(gc.fullname)='RE'
                                    ORDER BY it.sortorder
                                """ % (usuario)

                cursor.execute(sql)
                results = cursor.fetchall()
                if results:
                    for notasmooc in results:
                        campo = materiaasignada.campo(notasmooc[1].upper())
                        if not campo:
                            print('revisar curso moodle - ', materiaasignada.materia.id , 'idcursomoodle -', materiaasignada.materia.idcursomoodle )
                        if type(notasmooc[0]) is Decimal:
                            if null_to_decimal(campo.valor) != float(notasmooc[0]):
                                actualizar_nota_planificacion(materiaasignada.id, notasmooc[1].upper(), notasmooc[0])
                                auditorianotas = AuditoriaNotas(evaluaciongenerica=campo, manual=False,
                                                                calificacion=notasmooc[0])
                                auditorianotas.save()
                                print('importacion exitosa - ', materiaasignada)

                        else:
                            if null_to_decimal(campo.valor) != float(0):
                                actualizar_nota_planificacion(materiaasignada.id, notasmooc[1].upper(), notasmooc[0])
                                auditorianotas = AuditoriaNotas(evaluaciongenerica=campo, manual=False, calificacion=0)
                                auditorianotas.save()
                                print('importacion exitosa - ', materiaasignada)

                else:
                    detallemodelo = DetalleModeloEvaluativo.objects.get(pk=123)
                    campo = materiaasignada.campo(detallemodelo.nombre)
                    actualizar_nota_planificacion(materiaasignada.id, detallemodelo.nombre, 0)
                    auditorianotas = AuditoriaNotas(evaluaciongenerica=campo, manual=False, calificacion=0)
                    auditorianotas.save()



            print('PROCESO FINALIZADO')

        except Exception as ex:
            msg = ex.__str__()

            textoerror = '{} Linea:{}'.format(str(ex), sys.exc_info()[-1].tb_lineno)
            print(textoerror)
            print(msg)

def calificacion_transversales_domingo_presenciales_2():
    with transaction.atomic():
        try:
            periodo = Periodo.objects.get(id=177)
            asignaturas = DetalleGrupoAsignatura.objects.values_list('asignatura_id', flat=True).filter(status=True, grupo_id=1)
            materiasinvestigacion = MateriaAsignada.objects.filter(status=True,
                                                              materia__asignatura_id__in=asignaturas,
                                                              matricula__nivel__periodo=periodo,
                                                              matricula__bloqueomatricula=False,
                                                              matricula__retiradomatricula=False, materia__status=True,
                                                              matricula__status=True,
                                                              matricula__inscripcion__carrera__modalidad__in=[1],
                                                              matricula__inscripcion__carrera_id__in=[150,146,143,158,110,112,151,153,111,139,138],
                                                              materia__modeloevaluativo_id=27)
            cursor = connections['db_moodle_virtual'].cursor()
            for materiaasignada in materiasinvestigacion:
                # guardo_nota = False
                usuario = materiaasignada.matricula.inscripcion.persona.usuario.username
                #SE NECESITA EL ID DE CURSO MOODLE
                sql = """
                                    SELECT ROUND(nota.finalgrade,2), UPPER(gc.fullname)
                                            FROM mooc_grade_grades nota
                                    INNER JOIN mooc_grade_items it ON nota.itemid=it.id AND courseid=2040 AND itemtype='category'
                                    INNER JOIN mooc_grade_categories gc ON gc.courseid=it.courseid AND gc.id=it.iteminstance AND gc.depth=2
                                    INNER JOIN mooc_user us ON nota.userid=us.id
                                    WHERE us.username = '%s' and not UPPER(gc.fullname)='RE'
                                    ORDER BY it.sortorder
                                """ % (usuario)

                cursor.execute(sql)
                results = cursor.fetchall()
                if results:
                    for notasmooc in results:
                        campo = materiaasignada.campo(notasmooc[1].upper())
                        if not campo:
                            print('revisar curso moodle - ', materiaasignada.materia.id , 'idcursomoodle -', materiaasignada.materia.idcursomoodle )
                        if type(notasmooc[0]) is Decimal:
                            if null_to_decimal(campo.valor) != float(notasmooc[0]):
                                actualizar_nota_planificacion(materiaasignada.id, notasmooc[1].upper(), notasmooc[0])
                                auditorianotas = AuditoriaNotas(evaluaciongenerica=campo, manual=False,
                                                                calificacion=notasmooc[0])
                                auditorianotas.save()
                                print('importacion exitosa - ', materiaasignada)

                        else:
                            if null_to_decimal(campo.valor) != float(0):
                                actualizar_nota_planificacion(materiaasignada.id, notasmooc[1].upper(), notasmooc[0])
                                auditorianotas = AuditoriaNotas(evaluaciongenerica=campo, manual=False, calificacion=0)
                                auditorianotas.save()
                                print('importacion exitosa - ', materiaasignada)

                else:
                    detallemodelo = DetalleModeloEvaluativo.objects.get(pk=123)
                    campo = materiaasignada.campo(detallemodelo.nombre)
                    actualizar_nota_planificacion(materiaasignada.id, detallemodelo.nombre, 0)
                    auditorianotas = AuditoriaNotas(evaluaciongenerica=campo, manual=False, calificacion=0)
                    auditorianotas.save()



            print('PROCESO FINALIZADO')

        except Exception as ex:
            msg = ex.__str__()

            textoerror = '{} Linea:{}'.format(str(ex), sys.exc_info()[-1].tb_lineno)
            print(textoerror)
            print(msg)


def cerrar_materias_admision():

    for materia in Materia.objects.filter(status=True, nivel__periodo_id=177, nivel_id__in=[1481,1482]):
        for asig in materia.materiaasignada_set.filter(status=True, retiramateria=False).order_by(
                'matricula__inscripcion__persona__apellido1', 'matricula__inscripcion__persona__apellido2',
                'matricula__inscripcion__persona__nombres'):
            asig.cierre_materia_asignada()
            print(asig)
        materia.cerrado = True
        materia.fechacierre = datetime.now().date()
        materia.save()
        print(materia)
    for matricula in Matricula.objects.filter(status=True, nivel__periodo_id=177, nivel_id__in=[1481,1482]):
        matricula.cerrada = True
        matricula.save()

    print('fin')




def importar_cerrar_materias(carrera):
    with transaction.atomic():
        try:
            periodo = Periodo.objects.get(id=177)
            materias = Materia.objects.filter(status=True, nivel__periodo=periodo, cerrado=False, asignaturamalla__malla__carrera_id=carrera, fin__lte=datetime.now().date()).exclude(modeloevaluativo_id=27)
            for materia in materias:
                materiasasignadas = MateriaAsignada.objects.filter(status=True, materia=materia, matricula__bloqueomatricula=False,
                                                              matricula__retiradomatricula=False, matricula__status=True, retiramateria=False)

                for materiaasignada in materiasasignadas:
                    # guardo_nota = False
                    notas_de_moodle = materiaasignada.materia.notas_de_moodle(materiaasignada.matricula.inscripcion.persona)
                    if notas_de_moodle:
                        for notasmooc in notas_de_moodle:
                            campo = materiaasignada.campo(notasmooc[1].upper())
                            if not campo:
                                print('revisar curso moodle - ', materiaasignada.materia.id, 'idcursomoodle -',
                                      materiaasignada.materia.idcursomoodle)
                                continue
                            if type(notasmooc[0]) is Decimal:
                                if null_to_decimal(campo.valor) != float(notasmooc[0]):
                                    actualizar_nota_planificacion(materiaasignada.id, notasmooc[1].upper(),
                                                                  notasmooc[0])
                                    auditorianotas = AuditoriaNotas(evaluaciongenerica=campo, manual=False,
                                                                    calificacion=notasmooc[0])
                                    auditorianotas.save()
                                    print('importacion exitosa - ', materiaasignada)

                            else:
                                if null_to_decimal(campo.valor) != float(0):
                                    actualizar_nota_planificacion(materiaasignada.id, notasmooc[1].upper(),
                                                                  notasmooc[0])
                                    auditorianotas = AuditoriaNotas(evaluaciongenerica=campo, manual=False,
                                                                    calificacion=0)
                                    auditorianotas.save()
                                    print('importacion exitosa - ', materiaasignada)

                    else:
                        for detallemodelo in materiaasignada.materia.modeloevaluativo.detallemodeloevaluativo_set.filter(
                                migrarmoodle=True):
                            campo = materiaasignada.campo(detallemodelo.nombre)
                            actualizar_nota_planificacion(materiaasignada.id, detallemodelo.nombre, 0)
                            auditorianotas = AuditoriaNotas(evaluaciongenerica=campo, manual=False, calificacion=0)
                            auditorianotas.save()

                for asig in materia.asignados_a_esta_materia():
                    asig.cerrado = True
                    asig.save(actualiza=False)
                    asig.actualiza_estado()
                    asig.cierre_materia_asignada()
                # for asig in materia.asignados_a_esta_materia():
                #     asig.cierre_materia_asignada()

                materia.cerrado = True
                materia.fechacierre = datetime.now().date()
                materia.save()

            print('PROCESO FINALIZADO')

        except Exception as ex:
            msg = ex.__str__()

            textoerror = '{} Linea:{}'.format(str(ex), sys.exc_info()[-1].tb_lineno)
            print(textoerror)
            print(msg)

def cerrar_materias_transversales():
    materias = Materia.objects.filter(status=True, nivel__periodo_id=177, cerrada=False,
                                      asignaturamalla__malla__carrera__coordinacion__in=[1, 2, 3, 4, 5], modeloevaluativo_id=27)
    for materia in materias:
        for asig in materia.asignados_a_esta_materia():
            asig.cerrado = True
            asig.save(actualiza=False)
            asig.actualiza_estado()
        for asig in materia.asignados_a_esta_materia():
            asig.cierre_materia_asignada()

        materia.cerrado = True
        materia.fechacierre = datetime.now().date()
        materia.save()


def coordinacion1():
    #             materias = Materia.objects.filter(status=True, nivel__periodo=periodo, cerrado=False, asignaturamalla__malla__carrera_id=carrera, fin__lte=datetime.now().date()).exclude(modeloevaluativo_id=27)
    for carrera in Materia.objects.filter(status=True, nivel__periodo_id=177, cerrado=False,
                                          asignaturamalla__malla__carrera__coordinacion=1).exclude(
        modeloevaluativo_id=27).values_list('asignaturamalla__malla__carrera_id', flat=True).distinct():
        importar_cerrar_materias(carrera)



def calificacion_transversales_en_linea():
    try:
        periodo = Periodo.objects.get(id=177)
        asignaturas = DetalleGrupoAsignatura.objects.values_list('asignatura_id', flat=True).filter(status=True,
                                                                                                    grupo_id__in=[1, 2,
                                                                                                                  3])

        materias = Materia.objects.filter(status=True, nivel__periodo_id=177,
                                          asignaturamalla__asignatura_id__in=asignaturas,
                                          modeloevaluativo_id=27, asignaturamalla__malla__carrera__coordinacion__in=[2])

        for materia in materias:
            idcursomoodle = materia.idcursomoodle
            materiasasignadas = MateriaAsignada.objects.filter(status=True,
                                                               matricula__nivel__periodo=periodo,
                                                               materia=materia,
                                                               matricula__bloqueomatricula=False,
                                                               matricula__retiradomatricula=False, materia__status=True,
                                                               matricula__status=True,
                                                               matricula__inscripcion__carrera__modalidad__in=[1, 2, 3])

            cursor = connections['moodle_db'].cursor()
            for materiaasignada in materiasasignadas:
                # guardo_nota = False
                usuario = materiaasignada.matricula.inscripcion.persona.usuario.username
                # SE NECESITA EL ID DE CURSO MOODLE
                sql = """
                                                    SELECT ROUND(nota.finalgrade,2), UPPER(gc.fullname)
                                                            FROM mooc_grade_grades nota
                                                    INNER JOIN mooc_grade_items it ON nota.itemid=it.id AND courseid=%s AND itemtype='category'
                                                    INNER JOIN mooc_grade_categories gc ON gc.courseid=it.courseid AND gc.id=it.iteminstance AND gc.depth=2
                                                    INNER JOIN mooc_user us ON nota.userid=us.id
                                                    WHERE us.username = '%s' and UPPER(gc.fullname)='RE'
                                                    ORDER BY it.sortorder
                                                """ % (str(idcursomoodle), usuario)

                cursor.execute(sql)
                results = cursor.fetchall()
                if results:
                    for notasmooc in results:
                        campo = materiaasignada.campo(notasmooc[1].upper())
                        if not campo:
                            print('revisar curso moodle - ', materiaasignada.materia.id, 'idcursomoodle -',
                                  materiaasignada.materia.idcursomoodle)
                            continue
                        if type(notasmooc[0]) is Decimal:
                            if null_to_decimal(campo.valor) != float(notasmooc[0]):
                                actualizar_nota_planificacion(materiaasignada.id, notasmooc[1].upper(), notasmooc[0])
                                auditorianotas = AuditoriaNotas(evaluaciongenerica=campo, manual=False,
                                                                calificacion=notasmooc[0])
                                auditorianotas.save()
                                print('importacion exitosa - ', materiaasignada)

                        else:
                            if null_to_decimal(campo.valor) != float(0):
                                actualizar_nota_planificacion(materiaasignada.id, notasmooc[1].upper(), notasmooc[0])
                                auditorianotas = AuditoriaNotas(evaluaciongenerica=campo, manual=False, calificacion=0)
                                auditorianotas.save()
                                print('importacion exitosa - ', materiaasignada)

                else:
                    detallemodelo = DetalleModeloEvaluativo.objects.get(pk=125)
                    campo = materiaasignada.campo(detallemodelo.nombre)
                    actualizar_nota_planificacion(materiaasignada.id, detallemodelo.nombre, 0)
                    auditorianotas = AuditoriaNotas(evaluaciongenerica=campo, manual=False, calificacion=0)
                    auditorianotas.save()

            for asig in materia.asignados_a_esta_materia():
                asig.cerrado = True
                asig.save(actualiza=False)
                asig.actualiza_estado()
                asig.cierre_materia_asignada()

            print('PROCESO FINALIZADO')






    except Exception as ex:
        msg = ex.__str__()

        textoerror = '{} Linea:{}'.format(str(ex), sys.exc_info()[-1].tb_lineno)
        print(textoerror)
        print(msg)

def eliminar_seguimientos():
    seguimientos = SeguimientoTutor.objects.filter(materia_id__in=[72489,70619,70620,70701])
    for seguimiento in seguimientos:
        seguimiento.delete()
        print('registro eliminado')
    print('FIN')


def reajuste_horario_clase_seguida_pregrado(periodo_id):
    print(f"Inicio del proceso de reajuste de horarios del periodo_id: {periodo_id}")
    tipoprofesores = (
        (1, u'TEORIA'),
        (2, u'PRACTICA'),
        (7, u'VIRTUAL'),
        (11, u'AUTOR 2'),
        (12, u'AUTOR 1'),
        (10, u'ORIENTACION'),
        (14, u'PROFESOR TUTOR'),
        # (8, u'TUTOR VIRTUAL'),
        # (5, u'AYUDANTÍA')
    )

    tipohorarios = (
        (2, u'CLASE VIRTUAL SINCRÓNICA'),
        (7, u'CLASE VIRTUAL ASINCRÓNICA'),
        (8, u'CLASE REFUERZO SINCRÓNICA'),
        (9, u'CLASE REFUERZO ASINCRÓNICA'),
    )
    listadias = DIAS_CHOICES
    materias = Materia.objects.filter(nivel__periodo_id=periodo_id).exclude(asignaturamalla__transversal=True)

    total = len(materias)
    con = 0
    for materia in materias:
        Clase.objects.filter(status=True, materia=materia).update(subirenlace=False)
        print(f"Materia: {materia.__str__()}")
        for lisdia in listadias:
            for tipoprofesor in tipoprofesores:
                print(f"Tipo de profesor: {tipoprofesor[1]}")
                for tipohorario in tipohorarios:
                    print(f"Tipo de horario: {tipohorario[1]}")
                    if Clase.objects.filter(materia=materia, dia=lisdia[0], tipoprofesor_id=tipoprofesor[0], tipohorario=tipohorario[0]):
                        listaclase = Clase.objects.filter(materia=materia, dia=lisdia[0], tipoprofesor_id=tipoprofesor[0], tipohorario=tipohorario[0]).order_by('-id')
                        clasetrue = Clase.objects.filter(materia=materia, dia=lisdia[0], tipoprofesor_id=tipoprofesor[0], tipohorario=tipohorario[0]).order_by('-turno__comienza')[0]
                        clasetrue.subirenlace = True
                        clasetrue.save()
                        for lclase in listaclase:
                            if lclase.id != clasetrue.id:
                                lclase.subirenlace = False
                                lclase.save()
        con += 1
        print(f"{str(con)} / {str(total)}")
    print(f"Finalizo el proceso de reajuste de horarios del periodo_id: {periodo_id}")
#
#
#reajuste_horario_clase_seguida_pregrado(224)

def CrearExamenFinal():
    #excluir materias transversales
    materias=Materia.objects.filter(pk=73026)
    #materias = Materia.objects.filter(status=True, nivel__periodo_id=224, asignaturamalla__malla__modalidad_id=3, modeloevaluativo_id=7)
    no_silabos = []
    numsemana=16
    semana=2
    detallemodelo_id = 37
    persona = Persona.objects.get(pk=1)
    for materia in materias:
        try:
            eSilabo = Silabo.objects.filter(materia=materia, status=True, codigoqr=True).last()
            if eSilabo is None:
                print(f"Materia no tiene silabo {materia.__str__()}")
                no_silabos.append(materia.pk)
            else:
                try:
                    eSilaboSemanal = SilaboSemanal.objects.get(silabo=eSilabo, numsemana=numsemana, semana=semana,
                                                               examen=True)
                    eSilaboSemanal.delete()
                except ObjectDoesNotExist:
                    eSilaboSemanal = SilaboSemanal(silabo=eSilabo,
                                                   numsemana=numsemana,
                                                   semana=semana,
                                                   fechainiciosemana=datetime(2024, 1, 8, 0, 0).date(),
                                                   fechafinciosemana=datetime(2024, 1, 21, 0, 0).date(),
                                                   objetivoaprendizaje='',
                                                   enfoque='',
                                                   recursos='',
                                                   evaluacion='',
                                                   estado=3,
                                                   estadocumplimiento=2,
                                                   examen=True
                                                   )
                    eSilaboSemanal.save()
                # instruccion = """La honestidad es el pilar fundamental de toda sociedad, por ello la UNEMI motiva el desarrollo de este valor en cada uno de sus estudiantes.&nbsp;<br><br>Lea cada una de las preguntas y responda.&nbsp;<br><br>Esta actividad representa el 60% de la nota final.&nbsp;<br><br><br>"""
                instruccion = """Instrucciones;<br><br>Antes de ingresar a realizar el examen el estudiante debe revisar y estudiar todo el material del curso correspondiente al primer parcial.&nbsp;<br><br>Inicie el examen en la plataforma en la fecha y hora indicada. (No habrá prórroga).&nbsp;<br><br>Para resolver el presente Examen, dispone de un solo intento.&nbsp;<br><br>El examen debe ser resuelto en el tiempo máximo indicado, verificar su tiempo de disponibilidad.&nbsp;<br><br>Cualquier acto de deshonestidad académica será considerado como una falta y motivo de la inmediata suspensión del examen.&nbsp;<br><br><br>"""
                recomendacion = """Recomendaciones;<br><br>El examen tendrá una puntuación máxima de 20 puntos.&nbsp;<br><br>Preste atención a la hora de inicio del examen.&nbsp;<br><br>Si comienza el intento tiempo después, ese tiempo se le descontará de la hora preestablecida.&nbsp;<br><br>No se puede prorrogar el horario del examen.&nbsp;<br><br>Leer cuidadosamente cada pregunta y seguir la instrucción de la misma.&nbsp;<br><br>Una vez finalizado el examen no olvide dar clic en terminar el intento.&nbsp;<br><br><br>"""
                nombretest = 'EXAMEN FINAL'
                fechadesde = datetime(2024, 1, 8, 0, 0, 1)
                horadesde = datetime(2024, 1, 8, 0, 0, 1).time()
                fechahasta = datetime(2024, 1, 21, 0, 0, 1)
                horahasta = datetime(2024, 1, 21, 23, 59, 59).time()
                vecesintento = 1
                tiempoduracion = 60
                calificar = True
                navegacion = 1
                migrado = True
                password = '123JKDS@4pl'
                try:
                    eTestSilaboSemanal = TestSilaboSemanal.objects.get(status=True, silabosemanal=eSilaboSemanal,
                                                                       detallemodelo_id=detallemodelo_id)

                    eTestSilaboSemanal.delete()
                except ObjectDoesNotExist:
                    eTestSilaboSemanal = TestSilaboSemanal(detallemodelo_id=detallemodelo_id,
                                                           silabosemanal=eSilaboSemanal,
                                                           tiporecurso_id=11,
                                                           estado_id=1,
                                                           )
                eTestSilaboSemanal.detallemodelo_id = detallemodelo_id
                eTestSilaboSemanal.tiporecurso_id = 11
                eTestSilaboSemanal.estado_id = 1
                eTestSilaboSemanal.instruccion = instruccion
                eTestSilaboSemanal.recomendacion = recomendacion
                eTestSilaboSemanal.fechadesde = fechadesde
                eTestSilaboSemanal.horadesde = horadesde
                eTestSilaboSemanal.fechahasta = fechahasta
                eTestSilaboSemanal.horahasta = horahasta
                eTestSilaboSemanal.vecesintento = vecesintento
                eTestSilaboSemanal.tiempoduracion = tiempoduracion
                eTestSilaboSemanal.calificar = calificar
                eTestSilaboSemanal.navegacion = navegacion
                eTestSilaboSemanal.migrado = migrado
                eTestSilaboSemanal.password = password
                eTestSilaboSemanal.nombretest = nombretest
                eTestSilaboSemanal.save()

                # value, msg = CrearExamenMoodle(eTestSilaboSemanal.id, persona)
                # if not value:
                #     raise NameError(msg)




        except Exception as ex:
            transaction.set_rollback(True)
            print(f"Error {ex.__str__()}")

    if len(no_silabos):
        print(f"Materias sin silabos: {no_silabos.__str__()}")



def CrearExamenFinalRecuperacion():
    # excluir materias transversales
    # materias = Materia.objects.filter(status=True, nivel__periodo_id=224, asignaturamalla__malla__modalidad_id=3,
    #                                   modeloevaluativo_id=7)
    materias = Materia.objects.filter(pk=71551)


    no_silabos = []
    numsemana = 16
    semana = 2
    detallemodelo_id = 39
    persona = Persona.objects.get(pk=1)
    for materia in materias:
        try:
            eSilabo = Silabo.objects.filter(materia=materia, status=True, codigoqr=True).last()
            if eSilabo is None:
                print(f"Materia no tiene silabo {materia.__str__()}")
                no_silabos.append(materia.pk)
            else:
                try:
                    eSilaboSemanal = SilaboSemanal.objects.get(silabo=eSilabo, numsemana=numsemana, semana=semana,
                                                               examen=True)

                except ObjectDoesNotExist:
                    eSilaboSemanal = SilaboSemanal(silabo=eSilabo,
                                                   numsemana=numsemana,
                                                   semana=semana,
                                                   fechainiciosemana=datetime(2024, 1, 8, 0, 0).date(),
                                                   fechafinciosemana=datetime(2024, 1, 21, 0, 0).date(),
                                                   objetivoaprendizaje='',
                                                   enfoque='',
                                                   recursos='',
                                                   evaluacion='',
                                                   estado=3,
                                                   estadocumplimiento=2,
                                                   examen=True
                                                   )
                    eSilaboSemanal.save()
                # instruccion = """La honestidad es el pilar fundamental de toda sociedad, por ello la UNEMI motiva el desarrollo de este valor en cada uno de sus estudiantes.&nbsp;<br><br>Lea cada una de las preguntas y responda.&nbsp;<br><br>Esta actividad representa el 60% de la nota final.&nbsp;<br><br><br>"""
                instruccion = """Antes de ingresar a realizar el examen el estudiante debe revisar y estudiar todo el material del curso correspondiente al primer parcial.&nbsp;<br><br>Inicie el examen en la plataforma en la fecha y hora indicada. (No habrá prórroga).&nbsp;<br><br>Para resolver el presente Examen, dispone de un solo intento.&nbsp;<br><br>El examen debe ser resuelto en el tiempo máximo indicado, verificar su tiempo de disponibilidad.&nbsp;<br><br>Cualquier acto de deshonestidad académica será considerado como una falta y motivo de la inmediata suspensión del examen.&nbsp;<br><br><br>"""
                recomendacion = """El examen tendrá una puntuación máxima de 20 puntos.&nbsp;<br><br>Preste atención a la hora de inicio del examen.&nbsp;<br><br>Si comienza el intento tiempo después, ese tiempo se le descontará de la hora preestablecida.&nbsp;<br><br>No se puede prorrogar el horario del examen.&nbsp;<br><br>Leer cuidadosamente cada pregunta y seguir la instrucción de la misma.&nbsp;<br><br>Una vez finalizado el examen no olvide dar clic en terminar el intento.&nbsp;<br><br><br>"""
                nombretest = 'EXAMEN FINAL'
                fechadesde = datetime(2024, 1, 8, 0, 0, 1)
                horadesde = datetime(2024, 1, 8, 0, 0, 1).time()
                fechahasta = datetime(2024, 1, 21, 0, 0, 1)
                horahasta = datetime(2024, 1, 21, 23, 59, 59).time()
                vecesintento = 1
                tiempoduracion = 60
                calificar = True
                navegacion = 1
                migrado = True
                password = '123JKDS@4pl'
                try:
                    eTestSilaboSemanal = TestSilaboSemanal.objects.get(status=True, silabosemanal=eSilaboSemanal,
                                                                       detallemodelo_id=detallemodelo_id)
                except ObjectDoesNotExist:
                    eTestSilaboSemanal = TestSilaboSemanal(detallemodelo_id=detallemodelo_id,
                                                           silabosemanal=eSilaboSemanal,
                                                           tiporecurso_id=11,
                                                           estado_id=1,
                                                           )
                eTestSilaboSemanal.detallemodelo_id = detallemodelo_id
                eTestSilaboSemanal.tiporecurso_id = 11
                eTestSilaboSemanal.estado_id = 1
                eTestSilaboSemanal.instruccion = instruccion
                eTestSilaboSemanal.recomendacion = recomendacion
                eTestSilaboSemanal.fechadesde = fechadesde
                eTestSilaboSemanal.horadesde = horadesde
                eTestSilaboSemanal.fechahasta = fechahasta
                eTestSilaboSemanal.horahasta = horahasta
                eTestSilaboSemanal.vecesintento = vecesintento
                eTestSilaboSemanal.tiempoduracion = tiempoduracion
                eTestSilaboSemanal.calificar = calificar
                eTestSilaboSemanal.navegacion = navegacion
                eTestSilaboSemanal.migrado = migrado
                eTestSilaboSemanal.password = password
                eTestSilaboSemanal.nombretest = nombretest
                eTestSilaboSemanal.save()

                # value, msg = CrearExamenMoodle(eTestSilaboSemanal.id, persona)
                # if not value:
                #     raise NameError(msg)




        except Exception as ex:
            transaction.set_rollback(True)
            print(f"Error {ex.__str__()}")

    if len(no_silabos):
        print(f"Materias sin silabos: {no_silabos.__str__()}")


def importar_calificaciones():
    with transaction.atomic():
        try:
            periodo = Periodo.objects.get(id=224)
            materias = Materia.objects.filter(status=True, nivel__periodo=periodo, cerrado=False, asignaturamalla__malla__modalidad_id=3).exclude(nivel__id__in=[1516, 1517])
            for materia in materias:
                materiasasignadas = MateriaAsignada.objects.filter(status=True, materia=materia, matricula__bloqueomatricula=False,
                                                              matricula__retiradomatricula=False, matricula__status=True, retiramateria=False)

                for materiaasignada in materiasasignadas:
                    # guardo_nota = False
                    notas_de_moodle = materiaasignada.materia.notas_de_moodle(materiaasignada.matricula.inscripcion.persona)
                    if notas_de_moodle:
                        for notasmooc in notas_de_moodle:
                            campo = materiaasignada.campo(notasmooc[1].upper())
                            if not campo:
                                print('revisar curso moodle - ', materiaasignada.materia.id, 'idcursomoodle -',
                                      materiaasignada.materia.idcursomoodle)
                                continue
                            if type(notasmooc[0]) is Decimal:
                                if null_to_decimal(campo.valor) != float(notasmooc[0]):
                                    actualizar_nota_planificacion(materiaasignada.id, notasmooc[1].upper(),
                                                                  notasmooc[0])
                                    auditorianotas = AuditoriaNotas(evaluaciongenerica=campo, manual=False,
                                                                    calificacion=notasmooc[0])
                                    auditorianotas.save()
                                    print('importacion exitosa - ', materiaasignada)

                            else:
                                if null_to_decimal(campo.valor) != float(0):
                                    actualizar_nota_planificacion(materiaasignada.id, notasmooc[1].upper(),
                                                                  notasmooc[0])
                                    auditorianotas = AuditoriaNotas(evaluaciongenerica=campo, manual=False,
                                                                    calificacion=0)
                                    auditorianotas.save()
                                    print('importacion exitosa - ', materiaasignada)

                    else:
                        for detallemodelo in materiaasignada.materia.modeloevaluativo.detallemodeloevaluativo_set.filter(
                                migrarmoodle=True):
                            campo = materiaasignada.campo(detallemodelo.nombre)
                            actualizar_nota_planificacion(materiaasignada.id, detallemodelo.nombre, 0)
                            auditorianotas = AuditoriaNotas(evaluaciongenerica=campo, manual=False, calificacion=0)
                            auditorianotas.save()

                    materiaasignada.actualiza_estado()




            print('PROCESO FINALIZADO')

        except Exception as ex:
            msg = ex.__str__()

            textoerror = '{} Linea:{}'.format(str(ex), sys.exc_info()[-1].tb_lineno)
            print(textoerror)
            print(msg)


def actualizar_modelo_evaluativo():
    try:
        periodo = Periodo.objects.get(id=224)
        materias = Materia.objects.filter(pk=73026)
        #materias = Materia.objects.filter(status=True, modeloevaluativo_id=7, nivel__periodo=periodo, asignaturamalla__malla__carrera__modalidad__in=[3]).exclude(nivel_id__in=[1516,1517])
        total = materias.count()
        cont = 1
        from django.db import connections

        cursor = connections['moodle_db'].cursor()

        for materia in materias:


            #################################################################################################################
            # AGREGAR SISTEMA DE CALIFICACION
            #################################################################################################################
            if materia.idcursomoodle:
                cursoid = materia.idcursomoodle
                modelonotas = materia.modeloevaluativo.detallemodeloevaluativo_set.filter(migrarmoodle=True)
                if modelonotas:
                    query = u"SELECT id FROM mooc_grade_categories WHERE parent is null and depth=1 and courseid= %s" % cursoid
                    cursor.execute(query)
                    row = cursor.fetchall()
                    padrenota = 0
                    fecha = int(time.mktime(datetime.now().date().timetuple()))
                    if not row:
                        query = u"INSERT INTO mooc_grade_categories(courseid, parent, depth, path, fullname, aggregation, keephigh, droplow, aggregateonlygraded, hidden, timecreated, timemodified) VALUES (%s, null, 1, E'', E'?', 13, 0, 0, 0, 0, %s, %s)" % (
                            cursoid, fecha, fecha)
                        cursor.execute(query)
                        query = u"SELECT id FROM mooc_grade_categories WHERE parent is null and depth=1 and courseid= %s" % cursoid
                        cursor.execute(query)
                        row = cursor.fetchall()
                        query = u"UPDATE mooc_grade_categories SET path='/%s/' WHERE id= %s" % (row[0][0], row[0][0])
                        cursor.execute(query)
                        padrenota = row[0][0]
                    else:
                        padrenota = row[0][0]
                    if padrenota > 0:
                        ordennota = 1
                        query = u"SELECT id FROM mooc_grade_items WHERE courseid=%s and itemtype='course' and iteminstance=%s" % (
                            cursoid, padrenota)
                        cursor.execute(query)
                        row = cursor.fetchall()
                        if not row:
                            query = u"INSERT INTO mooc_grade_items (courseid, categoryid, itemname, itemtype, itemmodule, iteminstance, itemnumber, iteminfo, idnumber, calculation, gradetype, grademax, grademin, scaleid, outcomeid, gradepass, multfactor, plusfactor, aggregationcoef, aggregationcoef2, sortorder, display, decimals, hidden, locked, locktime, needsupdate, weightoverride, timecreated, timemodified) VALUES (%s, null, null, E'course', null, %s, null, null, null, null, 1, 100, 0, null, null, 0, 1, 0, 0, 0, %s, 0, 2, 0, 0, 0, 0, 0, %s, %s)" % (
                                cursoid, padrenota, ordennota, fecha, fecha)
                            cursor.execute(query)

                        for modelo in modelonotas:
                            query = u"SELECT id FROM mooc_grade_categories WHERE parent=%s and depth=2 and courseid= %s and fullname='%s'" % (
                                padrenota, cursoid, modelo.nombre)
                            cursor.execute(query)
                            row = cursor.fetchall()
                            padremodelo = 0
                            if not row:
                                query = u"INSERT INTO mooc_grade_categories(courseid, parent, depth, path, fullname, aggregation, keephigh, droplow, aggregateonlygraded, hidden, timecreated, timemodified) VALUES (%s, %s, 2, E'', E'%s', 0, 0, 0, 0, 0, %s, %s)" % (
                                    cursoid, padrenota, modelo.nombre, fecha, fecha)
                                cursor.execute(query)
                                query = u"SELECT id FROM mooc_grade_categories WHERE parent=%s and depth=2 and courseid= %s and fullname='%s'" % (
                                    padrenota, cursoid, modelo.nombre)
                                cursor.execute(query)
                                row = cursor.fetchall()
                                padremodelo = row[0][0]
                                query = u"UPDATE mooc_grade_categories SET path='/%s/%s/' WHERE id= %s" % (
                                    padrenota, padremodelo, padremodelo)
                                cursor.execute(query)
                            else:
                                padremodelo = row[0][0]
                            if padremodelo > 0:
                                ordennota += 1
                                query = u"SELECT id FROM mooc_grade_items WHERE courseid=%s and itemtype='category' and iteminstance=%s" % (
                                    cursoid, padremodelo)
                                cursor.execute(query)
                                row = cursor.fetchall()

                                if not row:
                                    if modelo.nombre == 'RE':
                                        query = u"INSERT INTO mooc_grade_items (courseid, categoryid, itemname, itemtype, itemmodule, iteminstance, itemnumber, iteminfo, idnumber, calculation, gradetype, grademax, grademin, scaleid, outcomeid, gradepass, multfactor, plusfactor, aggregationcoef, aggregationcoef2, sortorder, display, decimals, hidden, locked, locktime, needsupdate, weightoverride, timecreated, timemodified) " \
                                                u"VALUES (%s, null, E'', E'category', null, %s, null, E'', E'', null, 1, 0, 0, null, null, 0, 1, 0, 0, 0, %s, 0, null, 0, 0, 0, 0, 0, %s, %s)" \
                                                % (cursoid, padremodelo, ordennota,
                                                   fecha, fecha)
                                        cursor.execute(query)
                                    else:
                                        query = u"INSERT INTO mooc_grade_items (courseid, categoryid, itemname, itemtype, itemmodule, iteminstance, itemnumber, iteminfo, idnumber, calculation, gradetype, grademax, grademin, scaleid, outcomeid, gradepass, multfactor, plusfactor, aggregationcoef, aggregationcoef2, sortorder, display, decimals, hidden, locked, locktime, needsupdate, weightoverride, timecreated, timemodified) " \
                                                u"VALUES (%s, null, E'', E'category', null, %s, null, E'', E'', null, 1, %s, 0, null, null, 0, 1, 0, 0, %s, %s, 0, %s, 0, 0, 0, 0, 0, %s, %s)" \
                                                % (cursoid, padremodelo, modelo.notamaxima,
                                                   null_to_decimal(modelo.notamaxima / 100, 2), ordennota, 0,
                                                   fecha, fecha)
                                        cursor.execute(query)
                                else:
                                    gradeitemsid = row[0][0]
                                    if modelo.nombre == 'RE':
                                        query = u"UPDATE mooc_grade_items SET grademax='%s', aggregationcoef2='%s', decimals='%s'   WHERE id= %s" % (
                                            0, 0, 0, gradeitemsid )
                                        cursor.execute(query)



            #materia.crear_actualizar_categoria_notas_curso()
            print(cont, "de", total )
            cont += 1

        print('PROCESO FINALIZADO')

    except Exception as ex:
        msg = ex.__str__()

        textoerror = '{} Linea:{}'.format(str(ex), sys.exc_info()[-1].tb_lineno)
        print(textoerror)
        print(msg)

def actualizar_modelo():
    try:
        periodo = Periodo.objects.get(id=224)
        materias = Materia.objects.filter(status=True, nivel__periodo=periodo, modeloevaluativo_id=7).exclude(nivel_id__in=[1516,1517])
        total = materias.count()
        cont = 1

        for materia in materias:
            materia.crear_actualizar_categoria_notas_curso()
            print(cont, "de", total )
            cont += 1

        print('PROCESO FINALIZADO')

    except Exception as ex:
        msg = ex.__str__()

        textoerror = '{} Linea:{}'.format(str(ex), sys.exc_info()[-1].tb_lineno)
        print(textoerror)
        print(msg)

def actualizar_estado():
    try:
        periodo = Periodo.objects.get(id=224)
        materias = Materia.objects.filter(status=True, modeloevaluativo_id=27, nivel__periodo=periodo, cerrado=False,
                                          asignaturamalla__malla__modalidad_id__in=[1, 2]).exclude(
            nivel__id__in=[1516, 1517])
        for materia in materias:
            materiasasignadas = MateriaAsignada.objects.filter(status=True, materia=materia,
                                                               matricula__bloqueomatricula=False,
                                                               matricula__retiradomatricula=False,
                                                               matricula__status=True, retiramateria=False)

            for materiaasignada in materiasasignadas:
                materiaasignada.actualiza_estado()

        print('PROCESO FINALIZADO')

    except Exception as ex:
        msg = ex.__str__()

        textoerror = '{} Linea:{}'.format(str(ex), sys.exc_info()[-1].tb_lineno)
        print(textoerror)
        print(msg)

def importar_transversales():
    with transaction.atomic():
        try:
            periodo = Periodo.objects.get(id=224)
            materias = Materia.objects.filter(status=True, modeloevaluativo_id=27 ,nivel__periodo=periodo, cerrado=False, asignaturamalla__malla__modalidad_id__in=[1,2]).exclude(nivel__id__in=[1516, 1517])
            for materia in materias:
                materiasasignadas = MateriaAsignada.objects.filter(status=True, materia=materia, matricula__bloqueomatricula=False,
                                                              matricula__retiradomatricula=False, matricula__status=True, retiramateria=False)

                for materiaasignada in materiasasignadas:
                    # guardo_nota = False
                    notas_de_moodle = materiaasignada.materia.notas_de_moodle(materiaasignada.matricula.inscripcion.persona)
                    if notas_de_moodle:
                        for notasmooc in notas_de_moodle:
                            campo = materiaasignada.campo(notasmooc[1].upper())
                            if not campo:
                                print('revisar curso moodle - ', materiaasignada.materia.id, 'idcursomoodle -',
                                      materiaasignada.materia.idcursomoodle)
                                continue
                            if type(notasmooc[0]) is Decimal:
                                if null_to_decimal(campo.valor) != float(notasmooc[0]):
                                    actualizar_nota_planificacion(materiaasignada.id, notasmooc[1].upper(),
                                                                  notasmooc[0])
                                    auditorianotas = AuditoriaNotas(evaluaciongenerica=campo, manual=False,
                                                                    calificacion=notasmooc[0])
                                    auditorianotas.save()
                                    print('importacion exitosa - ', materiaasignada)

                            else:
                                if null_to_decimal(campo.valor) != float(0):
                                    actualizar_nota_planificacion(materiaasignada.id, notasmooc[1].upper(),
                                                                  notasmooc[0])
                                    auditorianotas = AuditoriaNotas(evaluaciongenerica=campo, manual=False,
                                                                    calificacion=0)
                                    auditorianotas.save()
                                    print('importacion exitosa - ', materiaasignada)

                    else:
                        for detallemodelo in materiaasignada.materia.modeloevaluativo.detallemodeloevaluativo_set.filter(
                                migrarmoodle=True):
                            campo = materiaasignada.campo(detallemodelo.nombre)
                            actualizar_nota_planificacion(materiaasignada.id, detallemodelo.nombre, 0)
                            auditorianotas = AuditoriaNotas(evaluaciongenerica=campo, manual=False, calificacion=0)
                            auditorianotas.save()






            print('PROCESO FINALIZADO')

        except Exception as ex:
            msg = ex.__str__()

            textoerror = '{} Linea:{}'.format(str(ex), sys.exc_info()[-1].tb_lineno)
            print(textoerror)
            print(msg)

def actualizar_estado():
    try:
        periodo = Periodo.objects.get(id=224)
        materias = Materia.objects.filter(status=True, modeloevaluativo_id=27, nivel__periodo=periodo, cerrado=False,
                                          asignaturamalla__malla__modalidad_id__in=[1, 2]).exclude(
            nivel__id__in=[1516, 1517])
        for materia in materias:
            materiasasignadas = MateriaAsignada.objects.filter(status=True, materia=materia,
                                                               matricula__bloqueomatricula=False,
                                                               matricula__retiradomatricula=False,
                                                               matricula__status=True, retiramateria=False)

            for materiaasignada in materiasasignadas:
                materiaasignada.actualiza_estado()

        print('PROCESO FINALIZADO')

    except Exception as ex:
        msg = ex.__str__()

        textoerror = '{} Linea:{}'.format(str(ex), sys.exc_info()[-1].tb_lineno)
        print(textoerror)
        print(msg)


@transaction.atomic()
def homologacion():

    #verificar que todos los estudiantes tengan la misma malla
    cadena = ''
    linea, excluidos, conexito = 0, 0, 0
    try:
        libre_origen = '/homologacion_derecho_2_1.xls'
        fuentecabecera = easyxf(
            'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
        fuentenormal = easyxf(
            'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
        output_folder = MEDIA_ROOT
        output_folder = os.path.join(os.path.join(BASE_DIR))
        # liborigen = xlrd.open_workbook(output_folder + libre_origen)
        libdestino = xlwt.Workbook()
        hojadestino = libdestino.add_sheet('HOJA1')
        fil = 0
        columnas = [(u"CEDULA", 6000),
                   (u"APELLIDOS Y NOMBRES", 6000),
                   (u"OBSERVACIÓN", 6000)
        ]
        for col_num in range(len(columnas)):
            hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
            hojadestino.col(col_num).width = columnas[col_num][1]
        fila = 1
        titulacion = 0

        lin = 0
        miarchivo = openpyxl.load_workbook("derecholinea.xlsx")
        #miarchivo = openpyxl.load_workbook("prueba_2.xlsx")

        ws = miarchivo.get_sheet_by_name("Hoja1")
        worksheet = ws
        c = 0
        cont = 0
        periodo_id=177
        carrera_id=126
        mallaantigua_id=207
        mallanueva_id=480

        for row in worksheet.iter_rows(min_row=0):
            if lin >= 0:
                currentValues, cadena = [], ''
                for cell in row:
                    cadena += str(cell.value) + ' '
                    currentValues.append(str(cell.value))
                identificacion = currentValues[0]

                matricula = Matricula.objects.filter(status=True, inscripcion__carrera__id=carrera_id, inscripcion__persona__cedula=identificacion).first()
                cont += 1
                matricula.pasoayuda = True
                matricula.save()
                print(u"%s - %s" % (matricula, cont))
                inscripcion = matricula.inscripcion
                hojadestino.write(fila, 0, matricula.inscripcion.persona.identificacion(), fuentenormal)
                hojadestino.write(fila, 1, matricula.inscripcion.persona.nombre_completo(), fuentenormal)
                hojadestino.write(fila, 2, matricula.nivel.periodo.nombre, fuentenormal)

                if InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallaantigua_id).exists():
                    imantigua = InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallaantigua_id)[0]
                    imantigua.status = False
                    imantigua.save()
                    print(u"Desactiva antigua inscripcion -----------------------------")

                if not InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallanueva_id).exists():
                    imnueva = InscripcionMalla(inscripcion=inscripcion, malla_id=mallanueva_id)
                    imnueva.save()
                    print(u"Crea nueva inscripcion -----------------------------")

                equivalencias = TablaEquivalenciaAsignaturas.objects.filter(status=True, asignaturamalla__malla_id=mallaantigua_id).order_by('asignaturamallasalto__nivelmalla__orden')
                for equivalencia in equivalencias:
                    print(u"nueva - %s" % equivalencia.asignaturamallasalto)
                    recordantiguo = inscripcion.recordacademico_set.filter(status=True,asignaturamalla_id=equivalencia.asignaturamalla_id).first()

                    if recordantiguo:
                        print(u"anterior - %s" % equivalencia.asignaturamalla)
                        print(u"Record antiguo: %s" % recordantiguo)
                        recordnuevo = None
                        recordantiguo.status = False
                        recordantiguo.save(update_asignaturamalla=False)

                        if equivalencia.asignaturamallasalto_id == 10623 or equivalencia.asignaturamallasalto_id == 10627 or equivalencia.asignaturamallasalto_id == 10615:
                            observaciones = recordantiguo.observaciones + " Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15"
                            homologada = True
                        else:
                            observaciones = recordantiguo.observaciones + " Migración con base a RESOLUCIÓN CGA-SO-9-2023-NO15 Y RESOLUCIÓN OCS-SO-18-2023-NO14"
                            homologada = False
                        if not RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                              asignaturamalla=equivalencia.asignaturamallasalto).exists():



                            recordnuevo = RecordAcademico(inscripcion=inscripcion,
                                                          matriculas=recordantiguo.matriculas,
                                                          asignaturamalla=equivalencia.asignaturamallasalto,
                                                          asignatura=equivalencia.asignaturamallasalto.asignatura,
                                                          asignaturaold_id=recordantiguo.asignatura.id,
                                                          nota=recordantiguo.nota,
                                                          asistencia=recordantiguo.asistencia,
                                                          sinasistencia=recordantiguo.sinasistencia,
                                                          fecha=recordantiguo.fecha,
                                                          noaplica=recordantiguo.noaplica,
                                                          aprobada=recordantiguo.aprobada,
                                                          convalidacion=recordantiguo.convalidacion,
                                                          pendiente=recordantiguo.pendiente,
                                                          creditos=equivalencia.asignaturamallasalto.creditos,
                                                          horas=equivalencia.asignaturamallasalto.horas,
                                                          valida=recordantiguo.valida,
                                                          validapromedio=recordantiguo.validapromedio,
                                                          observaciones=observaciones,
                                                          homologada=homologada,
                                                          materiaregular=recordantiguo.materiaregular,
                                                          materiacurso=None,
                                                          completonota=recordantiguo.completonota,
                                                          completoasistencia=recordantiguo.completoasistencia,
                                                          fechainicio=recordantiguo.fechainicio,
                                                          fechafin=recordantiguo.fechafin,
                                                          suficiencia=recordantiguo.suficiencia,
                                                          asignaturamallahistorico_id=recordantiguo.asignaturamalla.id,
                                                          reverso=False)
                            recordnuevo.save()
                            print(u"Crea nuevo record %s" % recordnuevo)


                        elif RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                            asignaturamalla=equivalencia.asignaturamallasalto):
                            recordnuevo = RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                         asignaturamalla=equivalencia.asignaturamallasalto)[0]
                            recordnuevo.matriculas = recordantiguo.matriculas
                            recordnuevo.asignaturamalla = equivalencia.asignaturamallasalto
                            recordnuevo.asignatura = equivalencia.asignaturamallasalto.asignatura
                            recordnuevo.asignaturaold = recordantiguo.asignaturamalla.asignatura
                            recordnuevo.nota = recordantiguo.nota
                            recordnuevo.asistencia = recordantiguo.asistencia
                            recordnuevo.sinasistencia = recordantiguo.sinasistencia
                            recordnuevo.fecha = recordantiguo.fecha
                            recordnuevo.noaplica = recordantiguo.noaplica
                            recordnuevo.aprobada = recordantiguo.aprobada
                            recordnuevo.convalidacion = recordantiguo.convalidacion
                            recordnuevo.pendiente = recordantiguo.pendiente
                            recordnuevo.creditos = equivalencia.asignaturamallasalto.creditos
                            recordnuevo.horas = equivalencia.asignaturamallasalto.horas
                            recordnuevo.valida = recordantiguo.valida
                            recordnuevo.validapromedio = recordantiguo.validapromedio
                            recordnuevo.observaciones = observaciones
                            recordnuevo.homologada = homologada
                            recordnuevo.materiaregular = recordantiguo.materiaregular
                            recordnuevo.materiacurso = None
                            recordnuevo.completonota = recordantiguo.completonota
                            recordnuevo.completoasistencia = recordantiguo.completoasistencia
                            recordnuevo.fechainicio = recordantiguo.fechainicio
                            recordnuevo.fechafin = recordantiguo.fechafin
                            recordnuevo.suficiencia = recordantiguo.suficiencia
                            recordnuevo.asignaturamallahistorico = recordantiguo.asignaturamalla
                            recordnuevo.reverso = False
                            recordnuevo.save()

                        if recordnuevo:
                            historicos = HistoricoRecordAcademico.objects.filter(status=True,
                                                                                 recordacademico=recordantiguo).update(recordacademico=recordnuevo,
                                                                                                                       creditos=recordnuevo.creditos,
                                                                                                                       horas=recordnuevo.horas,
                                                                                                                       homologada=recordnuevo.homologada)
                            respaldo = RespaldoRecordAcademico.objects.filter(status=True,recordacademicooriginal=recordantiguo)

                            if not respaldo.exists():
                                respaldorecord = RespaldoRecordAcademico(
                                    recordacademicooriginal=recordantiguo,
                                    recordacademiconuevo=recordnuevo
                                )
                                respaldorecord.save()
                            else:
                                respaldorecord = respaldo[0]
                                respaldorecord.recordacademiconuevo = recordnuevo
                                respaldorecord.save()
                            print(u"Record actualizado %s" % recordnuevo)


                    else:
                        hojadestino.write(fila, 3, "NO ENCONTRO RECORD ANTIGUO %s" % equivalencia.asignaturamalla)
                        fila += 1

                fila += 1

                time.sleep(3)

            lin += 1

        libdestino.save(output_folder + libre_origen)
        print(output_folder + libre_origen)
        print("Proceso finalizado. . .")


    except Exception as ex:
        noti = Notificacion(titulo='Error',
                            cuerpo='Ha ocurrido un error {} - Error en la linea {}'.format(ex, sys.exc_info()[
                                -1].tb_lineno),
                            destinatario_id=29898, url="",
                            prioridad=1, app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
                            tipo=2, en_proceso=False, error=True)
        noti.save()
        transaction.set_rollback(True)
        print('error: %s' % ex)
        hojadestino.write(fila, 3, str(ex))
        fila += 1

def homologacion_turismo():
    try:
        libre_origen = '/homologacion_turismo_2_1.xls'
        fuentecabecera = easyxf(
            'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
        fuentenormal = easyxf(
            'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')

        output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
        # liborigen = xlrd.open_workbook(output_folder + libre_origen)
        libdestino = xlwt.Workbook()
        hojadestino = libdestino.add_sheet('Sheet1')
        fil = 0
        columnas = [(u"CEDULA", 6000),
                    (u"APELLIDOS Y NOMBRES", 6000),
                    (u"OBSERVACIÓN", 6000)
                    ]
        for col_num in range(len(columnas)):
            hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
            hojadestino.col(col_num).width = columnas[col_num][1]
        fila = 1
        titulacion = 0

        lin = 0
        # miarchivo = openpyxl.load_workbook("CEDULA_COMUNICACION_final.xlsx")
        miarchivo = openpyxl.load_workbook("turismolinea.xlsx")
        # miarchivo = openpyxl.load_workbook(path_anexo)
        # miarchivo = openpyxl.load_workbook("prueba_2.xlsx")

        ws = miarchivo.get_sheet_by_name("primero")
        worksheet = ws
        c = 0
        cont = 0
        periodo_id = 224
        carrera_id = 134
        mallaantigua_id = 199
        mallanueva_id = 487
        for row in worksheet.iter_rows(min_row=0):
            if lin >= 0:
                currentValues, cadena = [], ''
                for cell in row:
                    cadena += str(cell.value) + ' '
                    currentValues.append(str(cell.value))
                identificacion = currentValues[0]

                if not identificacion:
                    break

                with transaction.atomic():
                    matricula = Matricula.objects.filter(status=True, inscripcion__carrera__id=carrera_id,
                                                         inscripcion__persona__cedula=identificacion).first()
                    cont += 1
                    matricula.pasoayuda = True
                    matricula.save()
                    print(u"%s - %s" % (matricula, cont))
                    inscripcion = matricula.inscripcion
                    hojadestino.write(fila, 0, matricula.inscripcion.persona.identificacion(), fuentenormal)
                    hojadestino.write(fila, 1, matricula.inscripcion.persona.nombre_completo(), fuentenormal)
                    hojadestino.write(fila, 2, matricula.nivel.periodo.nombre, fuentenormal)

                    itinerarios = ItinerariosMalla.objects.filter(status=True, malla_id=mallaantigua_id)

                    for itinerario in itinerarios:
                        itinerarionuevo = ItinerariosMalla.objects.filter(status=True, malla_id=mallanueva_id,
                                                                          nivel_id=itinerario.nivel_id)
                        # estadodo culmiado, en curso y pendiente
                        practicasencurso = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                               inscripcion=inscripcion,
                                                                                               estadosolicitud=2,
                                                                                               culminada=False,
                                                                                               itinerariomalla=itinerario)
                        for pcurso in practicasencurso:
                            pcurso.culminada = True
                            pcurso.save()

                        practicaspendientes = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                  inscripcion=inscripcion,
                                                                                                  estadosolicitud=4,
                                                                                                  culminada=False,
                                                                                                  itinerariomalla=itinerario)
                        for pendiente in practicaspendientes:
                            pendiente.estadosolicitud = 2
                            pendiente.culminada = True
                            pendiente.save()

                    practicaspp = haber_cumplido_horas_creditos_practicas_preprofesionales(inscripcion.id)
                    horasvinculacion = haber_cumplido_horas_creditos_vinculacion(inscripcion.id)

                    if InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                       malla_id=mallaantigua_id).exists():
                        imantigua = \
                            InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                            malla_id=mallaantigua_id)[
                                0]
                        imantigua.status = False
                        imantigua.save()
                        print(u"Desactiva antigua inscripcion -----------------------------")

                    if not InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                           malla_id=mallanueva_id).exists():
                        imnueva = InscripcionMalla(inscripcion=inscripcion, malla_id=mallanueva_id)
                        imnueva.save()
                        print(u"Crea nueva inscripcion -----------------------------")

                    equivalencias = TablaEquivalenciaAsignaturas.objects.filter(status=True,
                                                                                asignaturamalla__malla_id=mallaantigua_id).order_by(
                        'asignaturamallasalto__nivelmalla__orden')
                    for equivalencia in equivalencias:
                        print(u"nueva - %s" % equivalencia.asignaturamallasalto)
                        recordantiguo = inscripcion.recordacademico_set.filter(status=True,
                                                                               asignaturamalla_id=equivalencia.asignaturamalla_id).first()

                        if recordantiguo:
                            print(u"anterior - %s" % equivalencia.asignaturamalla)
                            print(u"Record antiguo: %s" % recordantiguo)
                            recordnuevo = None
                            recordantiguo.status = False
                            recordantiguo.save(update_asignaturamalla=False)

                            if equivalencia.asignaturamallasalto_id in [10929, 10947, 10959, 10953, 10964]:
                                observaciones = recordantiguo.observaciones + " Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15"
                                homologada = True



                            else:
                                observaciones = recordantiguo.observaciones + " Migración con base a RESOLUCIÓN CGA-SO-9-2023-NO15 Y RESOLUCIÓN OCS-SO-18-2023-NO14"
                                homologada = False
                            if not RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                  asignaturamalla=equivalencia.asignaturamallasalto).exists():

                                recordnuevo = RecordAcademico(inscripcion=inscripcion,
                                                              matriculas=recordantiguo.matriculas,
                                                              asignaturamalla=equivalencia.asignaturamallasalto,
                                                              asignatura=equivalencia.asignaturamallasalto.asignatura,
                                                              asignaturaold_id=recordantiguo.asignatura.id,
                                                              nota=recordantiguo.nota,
                                                              asistencia=recordantiguo.asistencia,
                                                              sinasistencia=recordantiguo.sinasistencia,
                                                              fecha=recordantiguo.fecha,
                                                              noaplica=recordantiguo.noaplica,
                                                              aprobada=recordantiguo.aprobada,
                                                              convalidacion=recordantiguo.convalidacion,
                                                              pendiente=recordantiguo.pendiente,
                                                              creditos=equivalencia.asignaturamallasalto.creditos,
                                                              horas=equivalencia.asignaturamallasalto.horas,
                                                              valida=recordantiguo.valida,
                                                              validapromedio=recordantiguo.validapromedio,
                                                              observaciones=observaciones,
                                                              homologada=homologada,
                                                              materiaregular=recordantiguo.materiaregular,
                                                              materiacurso=None,
                                                              completonota=recordantiguo.completonota,
                                                              completoasistencia=recordantiguo.completoasistencia,
                                                              fechainicio=recordantiguo.fechainicio,
                                                              fechafin=recordantiguo.fechafin,
                                                              suficiencia=recordantiguo.suficiencia,
                                                              asignaturamallahistorico_id=recordantiguo.asignaturamalla.id,
                                                              reverso=False)
                                recordnuevo.save()
                                print(u"Crea nuevo record %s" % recordnuevo)


                            elif RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                asignaturamalla=equivalencia.asignaturamallasalto):
                                recordnuevo = RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                             asignaturamalla=equivalencia.asignaturamallasalto)[
                                    0]
                                recordnuevo.matriculas = recordantiguo.matriculas
                                recordnuevo.asignaturamalla = equivalencia.asignaturamallasalto
                                recordnuevo.asignatura = equivalencia.asignaturamallasalto.asignatura
                                recordnuevo.asignaturaold = recordantiguo.asignaturamalla.asignatura
                                recordnuevo.nota = recordantiguo.nota
                                recordnuevo.asistencia = recordantiguo.asistencia
                                recordnuevo.sinasistencia = recordantiguo.sinasistencia
                                recordnuevo.fecha = recordantiguo.fecha
                                recordnuevo.noaplica = recordantiguo.noaplica
                                recordnuevo.aprobada = recordantiguo.aprobada
                                recordnuevo.convalidacion = recordantiguo.convalidacion
                                recordnuevo.pendiente = recordantiguo.pendiente
                                recordnuevo.creditos = equivalencia.asignaturamallasalto.creditos
                                recordnuevo.horas = equivalencia.asignaturamallasalto.horas
                                recordnuevo.valida = recordantiguo.valida
                                recordnuevo.validapromedio = recordantiguo.validapromedio
                                recordnuevo.observaciones = observaciones
                                recordnuevo.homologada = homologada
                                recordnuevo.materiaregular = recordantiguo.materiaregular
                                recordnuevo.materiacurso = None
                                recordnuevo.completonota = recordantiguo.completonota
                                recordnuevo.completoasistencia = recordantiguo.completoasistencia
                                recordnuevo.fechainicio = recordantiguo.fechainicio
                                recordnuevo.fechafin = recordantiguo.fechafin
                                recordnuevo.suficiencia = recordantiguo.suficiencia
                                recordnuevo.asignaturamallahistorico = recordantiguo.asignaturamalla
                                recordnuevo.reverso = False
                                recordnuevo.save()

                            if recordnuevo:
                                historicos = HistoricoRecordAcademico.objects.filter(status=True,
                                                                                     recordacademico=recordantiguo).update(
                                    recordacademico=recordnuevo,
                                    creditos=recordnuevo.creditos,
                                    horas=recordnuevo.horas,
                                    homologada=recordnuevo.homologada)
                                respaldo = RespaldoRecordAcademico.objects.filter(status=True,
                                                                                  recordacademicooriginal=recordantiguo)

                                if equivalencia.asignaturamallasalto_id in [10929, 10947, 10953]:
                                    if not practicaspp:
                                        if recordnuevo.aprobada:
                                            profesor = None
                                            if recordnuevo.materiaregular:
                                                profesor = recordnuevo.materiaregular.profesor_principal()
                                            elif recordnuevo.materiacurso:
                                                profesor = recordnuevo.materiaregular.profesor()
                                            if equivalencia.asignaturamallasalto_id == 10929:
                                                itinerarioquinto = ItinerariosMalla.objects.get(status=True,
                                                                                                malla_id=mallaantigua_id,
                                                                                                nivel_id=5)
                                                iquintonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                            malla_id=mallanueva_id,
                                                                                            nivel_id=5)

                                                practica = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud__in=[
                                                        1, 2, 4,
                                                        5, 6],
                                                    itinerariomalla=itinerarioquinto).exists()
                                                practicarechazada = False

                                                if not practica:
                                                    practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        estadosolicitud=3,
                                                        itinerariomalla=itinerarioquinto).exists()

                                                if not practica or practicarechazada:
                                                    if not PracticasPreprofesionalesInscripcion.objects.filter(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            actividad__itinerariomalla=itinerarioquinto).exists():
                                                        nuevapractica = PracticasPreprofesionalesInscripcion(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            numerohora=iquintonuevo.horas_practicas,
                                                            nivelmalla=iquintonuevo.nivel,
                                                            tiposolicitud=1,
                                                            estadosolicitud=2,
                                                            tipo=1,
                                                            itinerariomalla=iquintonuevo,
                                                            supervisor=profesor,
                                                            tutorunemi=profesor,
                                                            fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            tipoinstitucion=1,
                                                            sectoreconomico=6,
                                                            empresaempleadora_id=3,
                                                            culminada=True,
                                                            fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            lugarpractica_id=2,
                                                            observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                            )
                                                        nuevapractica.save()

                                                ItinerariosActividadDetalleDistributivoCarrera.objects.filter(
                                                    status=True,
                                                    itinerario=itinerarioquinto).update(
                                                    itinerario=iquintonuevo)


                                            if equivalencia.asignaturamallasalto_id == 10947:
                                                itinerariosexto = ItinerariosMalla.objects.get(status=True,
                                                                                               malla_id=mallaantigua_id,
                                                                                               nivel_id=6)
                                                isextonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallanueva_id,
                                                                                           nivel_id=6)

                                                practica = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud__in=[
                                                        1, 2, 4,
                                                        5, 6],
                                                    itinerariomalla=itinerariosexto).exists()
                                                practicarechazada = False

                                                if not practica:
                                                    practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        estadosolicitud=3,
                                                        itinerariomalla=itinerariosexto).exists()


                                                if not practica or practicarechazada:
                                                    if not PracticasPreprofesionalesInscripcion.objects.filter(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            actividad__itinerariomalla=itinerariosexto).exists():
                                                        nuevapractica = PracticasPreprofesionalesInscripcion(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                            fechahasta=recordnuevo.materiaregular.nivel.periodo.fin,
                                                            numerohora=isextonuevo.horas_practicas,
                                                            nivelmalla=isextonuevo.nivel,
                                                            tiposolicitud=1,
                                                            estadosolicitud=2,
                                                            tipo=1,
                                                            itinerariomalla=isextonuevo,
                                                            supervisor=profesor,
                                                            tutorunemi=profesor,
                                                            fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                            tipoinstitucion=1,
                                                            sectoreconomico=6,
                                                            empresaempleadora_id=3,
                                                            culminada=True,
                                                            fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                            lugarpractica_id=2,
                                                            observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                            )
                                                        nuevapractica.save()

                                                ItinerariosActividadDetalleDistributivoCarrera.objects.filter(
                                                    status=True,
                                                    itinerario=itinerariosexto).update(
                                                    itinerario=isextonuevo)


                                            if equivalencia.asignaturamallasalto_id == 10953:
                                                itinerarioseptimo = ItinerariosMalla.objects.get(status=True,
                                                                                                 malla_id=mallaantigua_id,
                                                                                                 nivel_id=7)
                                                iseptimonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                             malla_id=mallanueva_id,
                                                                                             nivel_id=7)

                                                practica = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud__in=[
                                                        1, 2, 4,
                                                        5, 6],
                                                    itinerariomalla=itinerarioseptimo).exists()
                                                practicarechazada = False

                                                if not practica:
                                                    practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        estadosolicitud=3,
                                                        itinerariomalla=itinerarioseptimo).exists()

                                                if not practica or practicarechazada:
                                                    if not PracticasPreprofesionalesInscripcion.objects.filter(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            actividad__itinerariomalla=itinerarioseptimo).exists():
                                                        nuevapractica = PracticasPreprofesionalesInscripcion(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                            fechahasta=recordnuevo.materiaregular.nivel.periodo.fin,
                                                            numerohora=iseptimonuevo.horas_practicas,
                                                            nivelmalla=iseptimonuevo.nivel,
                                                            tiposolicitud=1,
                                                            estadosolicitud=2,
                                                            tipo=1,
                                                            itinerariomalla=iseptimonuevo,
                                                            supervisor=profesor,
                                                            tutorunemi=profesor,
                                                            fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                            tipoinstitucion=1,
                                                            sectoreconomico=6,
                                                            empresaempleadora_id=3,
                                                            culminada=True,
                                                            fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                            lugarpractica_id=2,
                                                            observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                            )
                                                        nuevapractica.save()


                                if equivalencia.asignaturamallasalto_id in [10959, 10964]:
                                    if not horasvinculacion:
                                        if recordnuevo.aprobada:
                                            totalhoras = inscripcion.numero_horas_proyectos_vinculacion()
                                            if totalhoras == 0:
                                                a = 0
                                                print('sin horas')
                                            if equivalencia.asignaturamallasalto_id == 10959 and inscripcion.numero_horas_proyectos_vinculacion() < 80:
                                                horasfalta = 80 - inscripcion.numero_horas_proyectos_vinculacion()
                                                vinculacion = ParticipantesMatrices(status=True,
                                                                                    matrizevidencia_id=2,
                                                                                    proyecto_id=601,
                                                                                    inscripcion=inscripcion,
                                                                                    horas=horasfalta,
                                                                                    registrohorasdesde=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                                                    registrohorashasta=recordnuevo.materiaregular.nivel.periodo.fin,
                                                                                    estado=1
                                                                                    )
                                                vinculacion.save()

                                            if equivalencia.asignaturamallasalto_id == 10964 and inscripcion.numero_horas_proyectos_vinculacion() < 160:
                                                horasfalta = 160 - inscripcion.numero_horas_proyectos_vinculacion()
                                                vinculacion = ParticipantesMatrices(status=True,
                                                                                    matrizevidencia_id=2,
                                                                                    proyecto_id=601,
                                                                                    inscripcion=inscripcion,
                                                                                    horas=horasfalta,
                                                                                    registrohorasdesde=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                                                    registrohorashasta=recordnuevo.materiaregular.nivel.periodo.fin,
                                                                                    estado=1
                                                                                    )
                                                vinculacion.save()

                                if not respaldo.exists():
                                    respaldorecord = RespaldoRecordAcademico(
                                        recordacademicooriginal=recordantiguo,
                                        recordacademiconuevo=recordnuevo
                                    )
                                    respaldorecord.save()
                                else:
                                    respaldorecord = respaldo[0]
                                    respaldorecord.recordacademiconuevo = recordnuevo
                                    respaldorecord.save()
                                print(u"Record actualizado %s" % recordnuevo)


                        else:
                            hojadestino.write(fila, 3, "NO ENCONTRO RECORD ANTIGUO %s" % equivalencia.asignaturamalla)
                            fila += 1

                    fila += 1

                    time.sleep(3)



            lin += 1

        libdestino.save(output_folder + libre_origen)
        print(output_folder + libre_origen)
        print("Proceso finalizado. . .")


    except Exception as ex:

        transaction.set_rollback(True)
        print('error: %s' % ex)
        hojadestino.write(fila, 3, str(ex))
        fila += 1


def homologacion_turismo_2():
    try:
        libre_origen = '/homologacion_turismo_2_2.xls'
        fuentecabecera = easyxf(
            'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
        fuentenormal = easyxf(
            'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')

        output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
        # liborigen = xlrd.open_workbook(output_folder + libre_origen)
        libdestino = xlwt.Workbook()
        hojadestino = libdestino.add_sheet('Sheet1')
        fil = 0
        columnas = [(u"CEDULA", 6000),
                    (u"APELLIDOS Y NOMBRES", 6000),
                    (u"OBSERVACIÓN", 6000)
                    ]
        for col_num in range(len(columnas)):
            hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
            hojadestino.col(col_num).width = columnas[col_num][1]
        fila = 1
        titulacion = 0

        lin = 0
        # miarchivo = openpyxl.load_workbook("CEDULA_COMUNICACION_final.xlsx")
        miarchivo = openpyxl.load_workbook("turismolinea.xlsx")
        # miarchivo = openpyxl.load_workbook(path_anexo)
        # miarchivo = openpyxl.load_workbook("prueba_2.xlsx")

        ws = miarchivo.get_sheet_by_name("segundo")
        worksheet = ws
        c = 0
        cont = 0
        periodo_id = 224
        carrera_id = 134
        mallaantigua_id = 199
        mallanueva_id = 487
        for row in worksheet.iter_rows(min_row=0):
            if lin >= 0:
                currentValues, cadena = [], ''
                for cell in row:
                    cadena += str(cell.value) + ' '
                    currentValues.append(str(cell.value))
                identificacion = currentValues[0]

                if not identificacion:
                    break

                with transaction.atomic():
                    matricula = Matricula.objects.filter(status=True, inscripcion__carrera__id=carrera_id,
                                                         inscripcion__persona__cedula=identificacion).first()
                    cont += 1
                    matricula.pasoayuda = True
                    matricula.save()
                    print(u"%s - %s" % (matricula, cont))
                    inscripcion = matricula.inscripcion
                    hojadestino.write(fila, 0, matricula.inscripcion.persona.identificacion(), fuentenormal)
                    hojadestino.write(fila, 1, matricula.inscripcion.persona.nombre_completo(), fuentenormal)
                    hojadestino.write(fila, 2, matricula.nivel.periodo.nombre, fuentenormal)

                    itinerarios = ItinerariosMalla.objects.filter(status=True, malla_id=mallaantigua_id)

                    for itinerario in itinerarios:
                        itinerarionuevo = ItinerariosMalla.objects.filter(status=True, malla_id=mallanueva_id,
                                                                          nivel_id=itinerario.nivel_id)
                        # estadodo culmiado, en curso y pendiente
                        practicasencurso = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                               inscripcion=inscripcion,
                                                                                               estadosolicitud=2,
                                                                                               culminada=False,
                                                                                               itinerariomalla=itinerario)
                        for pcurso in practicasencurso:
                            pcurso.culminada = True
                            pcurso.save()

                        practicaspendientes = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                  inscripcion=inscripcion,
                                                                                                  estadosolicitud=4,
                                                                                                  culminada=False,
                                                                                                  itinerariomalla=itinerario)
                        for pendiente in practicaspendientes:
                            pendiente.estadosolicitud = 2
                            pendiente.culminada = True
                            pendiente.save()

                    practicaspp = haber_cumplido_horas_creditos_practicas_preprofesionales(inscripcion.id)
                    horasvinculacion = haber_cumplido_horas_creditos_vinculacion(inscripcion.id)

                    if InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                       malla_id=mallaantigua_id).exists():
                        imantigua = \
                            InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                            malla_id=mallaantigua_id)[
                                0]
                        imantigua.status = False
                        imantigua.save()
                        print(u"Desactiva antigua inscripcion -----------------------------")

                    if not InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                           malla_id=mallanueva_id).exists():
                        imnueva = InscripcionMalla(inscripcion=inscripcion, malla_id=mallanueva_id)
                        imnueva.save()
                        print(u"Crea nueva inscripcion -----------------------------")

                    equivalencias = TablaEquivalenciaAsignaturas.objects.filter(status=True,
                                                                                asignaturamalla__malla_id=mallaantigua_id).order_by(
                        'asignaturamallasalto__nivelmalla__orden')
                    for equivalencia in equivalencias:
                        print(u"nueva - %s" % equivalencia.asignaturamallasalto)
                        recordantiguo = inscripcion.recordacademico_set.filter(status=True,
                                                                               asignaturamalla_id=equivalencia.asignaturamalla_id).first()

                        if recordantiguo:
                            print(u"anterior - %s" % equivalencia.asignaturamalla)
                            print(u"Record antiguo: %s" % recordantiguo)
                            recordnuevo = None
                            recordantiguo.status = False
                            recordantiguo.save(update_asignaturamalla=False)

                            if equivalencia.asignaturamallasalto_id in [10929, 10947, 10959, 10953, 10964]:
                                observaciones = recordantiguo.observaciones + " Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15"
                                homologada = True



                            else:
                                observaciones = recordantiguo.observaciones + " Migración con base a RESOLUCIÓN CGA-SO-9-2023-NO15 Y RESOLUCIÓN OCS-SO-18-2023-NO14"
                                homologada = False
                            if not RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                  asignaturamalla=equivalencia.asignaturamallasalto).exists():

                                recordnuevo = RecordAcademico(inscripcion=inscripcion,
                                                              matriculas=recordantiguo.matriculas,
                                                              asignaturamalla=equivalencia.asignaturamallasalto,
                                                              asignatura=equivalencia.asignaturamallasalto.asignatura,
                                                              asignaturaold_id=recordantiguo.asignatura.id,
                                                              nota=recordantiguo.nota,
                                                              asistencia=recordantiguo.asistencia,
                                                              sinasistencia=recordantiguo.sinasistencia,
                                                              fecha=recordantiguo.fecha,
                                                              noaplica=recordantiguo.noaplica,
                                                              aprobada=recordantiguo.aprobada,
                                                              convalidacion=recordantiguo.convalidacion,
                                                              pendiente=recordantiguo.pendiente,
                                                              creditos=equivalencia.asignaturamallasalto.creditos,
                                                              horas=equivalencia.asignaturamallasalto.horas,
                                                              valida=recordantiguo.valida,
                                                              validapromedio=recordantiguo.validapromedio,
                                                              observaciones=observaciones,
                                                              homologada=homologada,
                                                              materiaregular=recordantiguo.materiaregular,
                                                              materiacurso=None,
                                                              completonota=recordantiguo.completonota,
                                                              completoasistencia=recordantiguo.completoasistencia,
                                                              fechainicio=recordantiguo.fechainicio,
                                                              fechafin=recordantiguo.fechafin,
                                                              suficiencia=recordantiguo.suficiencia,
                                                              asignaturamallahistorico_id=recordantiguo.asignaturamalla.id,
                                                              reverso=False)
                                recordnuevo.save()
                                print(u"Crea nuevo record %s" % recordnuevo)


                            elif RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                asignaturamalla=equivalencia.asignaturamallasalto):
                                recordnuevo = RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                             asignaturamalla=equivalencia.asignaturamallasalto)[
                                    0]
                                recordnuevo.matriculas = recordantiguo.matriculas
                                recordnuevo.asignaturamalla = equivalencia.asignaturamallasalto
                                recordnuevo.asignatura = equivalencia.asignaturamallasalto.asignatura
                                recordnuevo.asignaturaold = recordantiguo.asignaturamalla.asignatura
                                recordnuevo.nota = recordantiguo.nota
                                recordnuevo.asistencia = recordantiguo.asistencia
                                recordnuevo.sinasistencia = recordantiguo.sinasistencia
                                recordnuevo.fecha = recordantiguo.fecha
                                recordnuevo.noaplica = recordantiguo.noaplica
                                recordnuevo.aprobada = recordantiguo.aprobada
                                recordnuevo.convalidacion = recordantiguo.convalidacion
                                recordnuevo.pendiente = recordantiguo.pendiente
                                recordnuevo.creditos = equivalencia.asignaturamallasalto.creditos
                                recordnuevo.horas = equivalencia.asignaturamallasalto.horas
                                recordnuevo.valida = recordantiguo.valida
                                recordnuevo.validapromedio = recordantiguo.validapromedio
                                recordnuevo.observaciones = observaciones
                                recordnuevo.homologada = homologada
                                recordnuevo.materiaregular = recordantiguo.materiaregular
                                recordnuevo.materiacurso = None
                                recordnuevo.completonota = recordantiguo.completonota
                                recordnuevo.completoasistencia = recordantiguo.completoasistencia
                                recordnuevo.fechainicio = recordantiguo.fechainicio
                                recordnuevo.fechafin = recordantiguo.fechafin
                                recordnuevo.suficiencia = recordantiguo.suficiencia
                                recordnuevo.asignaturamallahistorico = recordantiguo.asignaturamalla
                                recordnuevo.reverso = False
                                recordnuevo.save()

                            if recordnuevo:
                                historicos = HistoricoRecordAcademico.objects.filter(status=True,
                                                                                     recordacademico=recordantiguo).update(
                                    recordacademico=recordnuevo,
                                    creditos=recordnuevo.creditos,
                                    horas=recordnuevo.horas,
                                    homologada=recordnuevo.homologada)
                                respaldo = RespaldoRecordAcademico.objects.filter(status=True,
                                                                                  recordacademicooriginal=recordantiguo)

                                if equivalencia.asignaturamallasalto_id in [10929, 10947, 10953]:
                                    if not practicaspp:
                                        if recordnuevo.aprobada:
                                            profesor = None
                                            if recordnuevo.materiaregular:
                                                profesor = recordnuevo.materiaregular.profesor_principal()
                                            elif recordnuevo.materiacurso:
                                                profesor = recordnuevo.materiaregular.profesor()
                                            if equivalencia.asignaturamallasalto_id == 10929:
                                                itinerarioquinto = ItinerariosMalla.objects.get(status=True,
                                                                                                malla_id=mallaantigua_id,
                                                                                                nivel_id=5)
                                                iquintonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                            malla_id=mallanueva_id,
                                                                                            nivel_id=5)

                                                practica = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud__in=[
                                                        1, 2, 4,
                                                        5, 6],
                                                    itinerariomalla=itinerarioquinto).exists()
                                                practicarechazada = False

                                                if not practica:
                                                    practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        estadosolicitud=3,
                                                        itinerariomalla=itinerarioquinto).exists()

                                                if not practica or practicarechazada:
                                                    if not PracticasPreprofesionalesInscripcion.objects.filter(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            actividad__itinerariomalla=itinerarioquinto).exists():
                                                        nuevapractica = PracticasPreprofesionalesInscripcion(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            numerohora=iquintonuevo.horas_practicas,
                                                            nivelmalla=iquintonuevo.nivel,
                                                            tiposolicitud=1,
                                                            estadosolicitud=2,
                                                            tipo=1,
                                                            itinerariomalla=iquintonuevo,
                                                            supervisor=profesor,
                                                            tutorunemi=profesor,
                                                            fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            tipoinstitucion=1,
                                                            sectoreconomico=6,
                                                            empresaempleadora_id=3,
                                                            culminada=True,
                                                            fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            lugarpractica_id=2,
                                                            observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                        )
                                                        nuevapractica.save()

                                                ItinerariosActividadDetalleDistributivoCarrera.objects.filter(
                                                    status=True,
                                                    itinerario=itinerarioquinto).update(
                                                    itinerario=iquintonuevo)

                                            if equivalencia.asignaturamallasalto_id == 10947:
                                                itinerariosexto = ItinerariosMalla.objects.get(status=True,
                                                                                               malla_id=mallaantigua_id,
                                                                                               nivel_id=6)
                                                isextonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallanueva_id,
                                                                                           nivel_id=6)

                                                practica = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud__in=[
                                                        1, 2, 4,
                                                        5, 6],
                                                    itinerariomalla=itinerariosexto).exists()
                                                practicarechazada = False

                                                if not practica:
                                                    practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        estadosolicitud=3,
                                                        itinerariomalla=itinerariosexto).exists()

                                                if not practica or practicarechazada:
                                                    if not PracticasPreprofesionalesInscripcion.objects.filter(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            actividad__itinerariomalla=itinerariosexto).exists():
                                                        nuevapractica = PracticasPreprofesionalesInscripcion(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                            fechahasta=recordnuevo.materiaregular.nivel.periodo.fin,
                                                            numerohora=isextonuevo.horas_practicas,
                                                            nivelmalla=isextonuevo.nivel,
                                                            tiposolicitud=1,
                                                            estadosolicitud=2,
                                                            tipo=1,
                                                            itinerariomalla=isextonuevo,
                                                            supervisor=profesor,
                                                            tutorunemi=profesor,
                                                            fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                            tipoinstitucion=1,
                                                            sectoreconomico=6,
                                                            empresaempleadora_id=3,
                                                            culminada=True,
                                                            fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                            lugarpractica_id=2,
                                                            observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                        )
                                                        nuevapractica.save()

                                                ItinerariosActividadDetalleDistributivoCarrera.objects.filter(
                                                    status=True,
                                                    itinerario=itinerariosexto).update(
                                                    itinerario=isextonuevo)

                                            if equivalencia.asignaturamallasalto_id == 10953:
                                                itinerarioseptimo = ItinerariosMalla.objects.get(status=True,
                                                                                                 malla_id=mallaantigua_id,
                                                                                                 nivel_id=7)
                                                iseptimonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                             malla_id=mallanueva_id,
                                                                                             nivel_id=7)

                                                practica = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud__in=[
                                                        1, 2, 4,
                                                        5, 6],
                                                    itinerariomalla=itinerarioseptimo).exists()
                                                practicarechazada = False

                                                if not practica:
                                                    practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        estadosolicitud=3,
                                                        itinerariomalla=itinerarioseptimo).exists()

                                                if not practica or practicarechazada:
                                                    if not PracticasPreprofesionalesInscripcion.objects.filter(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            actividad__itinerariomalla=itinerarioseptimo).exists():
                                                        nuevapractica = PracticasPreprofesionalesInscripcion(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                            fechahasta=recordnuevo.materiaregular.nivel.periodo.fin,
                                                            numerohora=iseptimonuevo.horas_practicas,
                                                            nivelmalla=iseptimonuevo.nivel,
                                                            tiposolicitud=1,
                                                            estadosolicitud=2,
                                                            tipo=1,
                                                            itinerariomalla=iseptimonuevo,
                                                            supervisor=profesor,
                                                            tutorunemi=profesor,
                                                            fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                            tipoinstitucion=1,
                                                            sectoreconomico=6,
                                                            empresaempleadora_id=3,
                                                            culminada=True,
                                                            fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                            lugarpractica_id=2,
                                                            observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                        )
                                                        nuevapractica.save()

                                if equivalencia.asignaturamallasalto_id in [10959, 10964]:
                                    if not horasvinculacion:
                                        if recordnuevo.aprobada:
                                            totalhoras = inscripcion.numero_horas_proyectos_vinculacion()
                                            if totalhoras == 0:
                                                a = 0
                                                print('sin horas')
                                            if equivalencia.asignaturamallasalto_id == 10959 and inscripcion.numero_horas_proyectos_vinculacion() < 80:
                                                horasfalta = 80 - inscripcion.numero_horas_proyectos_vinculacion()
                                                vinculacion = ParticipantesMatrices(status=True,
                                                                                    matrizevidencia_id=2,
                                                                                    proyecto_id=601,
                                                                                    inscripcion=inscripcion,
                                                                                    horas=horasfalta,
                                                                                    registrohorasdesde=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                                                    registrohorashasta=recordnuevo.materiaregular.nivel.periodo.fin,
                                                                                    estado=1
                                                                                    )
                                                vinculacion.save()

                                            if equivalencia.asignaturamallasalto_id == 10964 and inscripcion.numero_horas_proyectos_vinculacion() < 160:
                                                horasfalta = 160 - inscripcion.numero_horas_proyectos_vinculacion()
                                                vinculacion = ParticipantesMatrices(status=True,
                                                                                    matrizevidencia_id=2,
                                                                                    proyecto_id=601,
                                                                                    inscripcion=inscripcion,
                                                                                    horas=horasfalta,
                                                                                    registrohorasdesde=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                                                    registrohorashasta=recordnuevo.materiaregular.nivel.periodo.fin,
                                                                                    estado=1
                                                                                    )
                                                vinculacion.save()

                                if not respaldo.exists():
                                    respaldorecord = RespaldoRecordAcademico(
                                        recordacademicooriginal=recordantiguo,
                                        recordacademiconuevo=recordnuevo
                                    )
                                    respaldorecord.save()
                                else:
                                    respaldorecord = respaldo[0]
                                    respaldorecord.recordacademiconuevo = recordnuevo
                                    respaldorecord.save()
                                print(u"Record actualizado %s" % recordnuevo)


                        else:
                            hojadestino.write(fila, 3, "NO ENCONTRO RECORD ANTIGUO %s" % equivalencia.asignaturamalla)
                            fila += 1

                    fila += 1

                    time.sleep(3)

            lin += 1

        libdestino.save(output_folder + libre_origen)
        print(output_folder + libre_origen)
        print("Proceso finalizado. . .")


    except Exception as ex:

        transaction.set_rollback(True)
        print('error: %s' % ex)
        hojadestino.write(fila, 3, str(ex))
        fila += 1

def homologacion_turismo_3():
    try:
        libre_origen = '/homologacion_turismo_2_3.xls'
        fuentecabecera = easyxf(
            'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
        fuentenormal = easyxf(
            'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')

        output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
        # liborigen = xlrd.open_workbook(output_folder + libre_origen)
        libdestino = xlwt.Workbook()
        hojadestino = libdestino.add_sheet('Sheet1')
        fil = 0
        columnas = [(u"CEDULA", 6000),
                    (u"APELLIDOS Y NOMBRES", 6000),
                    (u"OBSERVACIÓN", 6000)
                    ]
        for col_num in range(len(columnas)):
            hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
            hojadestino.col(col_num).width = columnas[col_num][1]
        fila = 1
        titulacion = 0

        lin = 0
        # miarchivo = openpyxl.load_workbook("CEDULA_COMUNICACION_final.xlsx")
        miarchivo = openpyxl.load_workbook("turismolinea.xlsx")
        # miarchivo = openpyxl.load_workbook(path_anexo)
        # miarchivo = openpyxl.load_workbook("prueba_2.xlsx")

        ws = miarchivo.get_sheet_by_name("tercero")
        worksheet = ws
        c = 0
        cont = 0
        periodo_id = 224
        carrera_id = 134
        mallaantigua_id = 199
        mallanueva_id = 487
        for row in worksheet.iter_rows(min_row=0):
            if lin >= 0:
                currentValues, cadena = [], ''
                for cell in row:
                    cadena += str(cell.value) + ' '
                    currentValues.append(str(cell.value))
                identificacion = currentValues[0]

                if not identificacion:
                    break

                with transaction.atomic():
                    matricula = Matricula.objects.filter(status=True, inscripcion__carrera__id=carrera_id,
                                                         inscripcion__persona__cedula=identificacion).first()
                    cont += 1
                    matricula.pasoayuda = True
                    matricula.save()
                    print(u"%s - %s" % (matricula, cont))
                    inscripcion = matricula.inscripcion
                    hojadestino.write(fila, 0, matricula.inscripcion.persona.identificacion(), fuentenormal)
                    hojadestino.write(fila, 1, matricula.inscripcion.persona.nombre_completo(), fuentenormal)
                    hojadestino.write(fila, 2, matricula.nivel.periodo.nombre, fuentenormal)

                    itinerarios = ItinerariosMalla.objects.filter(status=True, malla_id=mallaantigua_id)

                    for itinerario in itinerarios:
                        itinerarionuevo = ItinerariosMalla.objects.filter(status=True, malla_id=mallanueva_id,
                                                                          nivel_id=itinerario.nivel_id)
                        # estadodo culmiado, en curso y pendiente
                        practicasencurso = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                               inscripcion=inscripcion,
                                                                                               estadosolicitud=2,
                                                                                               culminada=False,
                                                                                               itinerariomalla=itinerario)
                        for pcurso in practicasencurso:
                            pcurso.culminada = True
                            pcurso.save()

                        practicaspendientes = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                  inscripcion=inscripcion,
                                                                                                  estadosolicitud=4,
                                                                                                  culminada=False,
                                                                                                  itinerariomalla=itinerario)
                        for pendiente in practicaspendientes:
                            pendiente.estadosolicitud = 2
                            pendiente.culminada = True
                            pendiente.save()

                    practicaspp = haber_cumplido_horas_creditos_practicas_preprofesionales(inscripcion.id)
                    horasvinculacion = haber_cumplido_horas_creditos_vinculacion(inscripcion.id)

                    if InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                       malla_id=mallaantigua_id).exists():
                        imantigua = \
                            InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                            malla_id=mallaantigua_id)[
                                0]
                        imantigua.status = False
                        imantigua.save()
                        print(u"Desactiva antigua inscripcion -----------------------------")

                    if not InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                           malla_id=mallanueva_id).exists():
                        imnueva = InscripcionMalla(inscripcion=inscripcion, malla_id=mallanueva_id)
                        imnueva.save()
                        print(u"Crea nueva inscripcion -----------------------------")

                    equivalencias = TablaEquivalenciaAsignaturas.objects.filter(status=True,
                                                                                asignaturamalla__malla_id=mallaantigua_id).order_by(
                        'asignaturamallasalto__nivelmalla__orden')
                    for equivalencia in equivalencias:
                        print(u"nueva - %s" % equivalencia.asignaturamallasalto)
                        recordantiguo = inscripcion.recordacademico_set.filter(status=True,
                                                                               asignaturamalla_id=equivalencia.asignaturamalla_id).first()

                        if recordantiguo:
                            print(u"anterior - %s" % equivalencia.asignaturamalla)
                            print(u"Record antiguo: %s" % recordantiguo)
                            recordnuevo = None
                            recordantiguo.status = False
                            recordantiguo.save(update_asignaturamalla=False)

                            if equivalencia.asignaturamallasalto_id in [10929, 10947, 10959, 10953, 10964]:
                                observaciones = recordantiguo.observaciones + " Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15"
                                homologada = True



                            else:
                                observaciones = recordantiguo.observaciones + " Migración con base a RESOLUCIÓN CGA-SO-9-2023-NO15 Y RESOLUCIÓN OCS-SO-18-2023-NO14"
                                homologada = False
                            if not RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                  asignaturamalla=equivalencia.asignaturamallasalto).exists():

                                recordnuevo = RecordAcademico(inscripcion=inscripcion,
                                                              matriculas=recordantiguo.matriculas,
                                                              asignaturamalla=equivalencia.asignaturamallasalto,
                                                              asignatura=equivalencia.asignaturamallasalto.asignatura,
                                                              asignaturaold_id=recordantiguo.asignatura.id,
                                                              nota=recordantiguo.nota,
                                                              asistencia=recordantiguo.asistencia,
                                                              sinasistencia=recordantiguo.sinasistencia,
                                                              fecha=recordantiguo.fecha,
                                                              noaplica=recordantiguo.noaplica,
                                                              aprobada=recordantiguo.aprobada,
                                                              convalidacion=recordantiguo.convalidacion,
                                                              pendiente=recordantiguo.pendiente,
                                                              creditos=equivalencia.asignaturamallasalto.creditos,
                                                              horas=equivalencia.asignaturamallasalto.horas,
                                                              valida=recordantiguo.valida,
                                                              validapromedio=recordantiguo.validapromedio,
                                                              observaciones=observaciones,
                                                              homologada=homologada,
                                                              materiaregular=recordantiguo.materiaregular,
                                                              materiacurso=None,
                                                              completonota=recordantiguo.completonota,
                                                              completoasistencia=recordantiguo.completoasistencia,
                                                              fechainicio=recordantiguo.fechainicio,
                                                              fechafin=recordantiguo.fechafin,
                                                              suficiencia=recordantiguo.suficiencia,
                                                              asignaturamallahistorico_id=recordantiguo.asignaturamalla.id,
                                                              reverso=False)
                                recordnuevo.save()
                                print(u"Crea nuevo record %s" % recordnuevo)


                            elif RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                asignaturamalla=equivalencia.asignaturamallasalto):
                                recordnuevo = RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                             asignaturamalla=equivalencia.asignaturamallasalto)[
                                    0]
                                recordnuevo.matriculas = recordantiguo.matriculas
                                recordnuevo.asignaturamalla = equivalencia.asignaturamallasalto
                                recordnuevo.asignatura = equivalencia.asignaturamallasalto.asignatura
                                recordnuevo.asignaturaold = recordantiguo.asignaturamalla.asignatura
                                recordnuevo.nota = recordantiguo.nota
                                recordnuevo.asistencia = recordantiguo.asistencia
                                recordnuevo.sinasistencia = recordantiguo.sinasistencia
                                recordnuevo.fecha = recordantiguo.fecha
                                recordnuevo.noaplica = recordantiguo.noaplica
                                recordnuevo.aprobada = recordantiguo.aprobada
                                recordnuevo.convalidacion = recordantiguo.convalidacion
                                recordnuevo.pendiente = recordantiguo.pendiente
                                recordnuevo.creditos = equivalencia.asignaturamallasalto.creditos
                                recordnuevo.horas = equivalencia.asignaturamallasalto.horas
                                recordnuevo.valida = recordantiguo.valida
                                recordnuevo.validapromedio = recordantiguo.validapromedio
                                recordnuevo.observaciones = observaciones
                                recordnuevo.homologada = homologada
                                recordnuevo.materiaregular = recordantiguo.materiaregular
                                recordnuevo.materiacurso = None
                                recordnuevo.completonota = recordantiguo.completonota
                                recordnuevo.completoasistencia = recordantiguo.completoasistencia
                                recordnuevo.fechainicio = recordantiguo.fechainicio
                                recordnuevo.fechafin = recordantiguo.fechafin
                                recordnuevo.suficiencia = recordantiguo.suficiencia
                                recordnuevo.asignaturamallahistorico = recordantiguo.asignaturamalla
                                recordnuevo.reverso = False
                                recordnuevo.save()

                            if recordnuevo:
                                historicos = HistoricoRecordAcademico.objects.filter(status=True,
                                                                                     recordacademico=recordantiguo).update(
                                    recordacademico=recordnuevo,
                                    creditos=recordnuevo.creditos,
                                    horas=recordnuevo.horas,
                                    homologada=recordnuevo.homologada)
                                respaldo = RespaldoRecordAcademico.objects.filter(status=True,
                                                                                  recordacademicooriginal=recordantiguo)

                                if equivalencia.asignaturamallasalto_id in [10929, 10947, 10953]:
                                    if not practicaspp:
                                        if recordnuevo.aprobada:
                                            profesor = None
                                            if recordnuevo.materiaregular:
                                                profesor = recordnuevo.materiaregular.profesor_principal()
                                            elif recordnuevo.materiacurso:
                                                profesor = recordnuevo.materiaregular.profesor()
                                            if equivalencia.asignaturamallasalto_id == 10929:
                                                itinerarioquinto = ItinerariosMalla.objects.get(status=True,
                                                                                                malla_id=mallaantigua_id,
                                                                                                nivel_id=5)
                                                iquintonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                            malla_id=mallanueva_id,
                                                                                            nivel_id=5)

                                                practica = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud__in=[
                                                        1, 2, 4,
                                                        5, 6],
                                                    itinerariomalla=itinerarioquinto).exists()
                                                practicarechazada = False

                                                if not practica:
                                                    practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        estadosolicitud=3,
                                                        itinerariomalla=itinerarioquinto).exists()

                                                if not practica or practicarechazada:
                                                    if not PracticasPreprofesionalesInscripcion.objects.filter(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            actividad__itinerariomalla=itinerarioquinto).exists():
                                                        nuevapractica = PracticasPreprofesionalesInscripcion(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            numerohora=iquintonuevo.horas_practicas,
                                                            nivelmalla=iquintonuevo.nivel,
                                                            tiposolicitud=1,
                                                            estadosolicitud=2,
                                                            tipo=1,
                                                            itinerariomalla=iquintonuevo,
                                                            supervisor=profesor,
                                                            tutorunemi=profesor,
                                                            fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            tipoinstitucion=1,
                                                            sectoreconomico=6,
                                                            empresaempleadora_id=3,
                                                            culminada=True,
                                                            fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            lugarpractica_id=2,
                                                            observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                        )
                                                        nuevapractica.save()

                                                ItinerariosActividadDetalleDistributivoCarrera.objects.filter(
                                                    status=True,
                                                    itinerario=itinerarioquinto).update(
                                                    itinerario=iquintonuevo)

                                            if equivalencia.asignaturamallasalto_id == 10947:
                                                itinerariosexto = ItinerariosMalla.objects.get(status=True,
                                                                                               malla_id=mallaantigua_id,
                                                                                               nivel_id=6)
                                                isextonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallanueva_id,
                                                                                           nivel_id=6)

                                                practica = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud__in=[
                                                        1, 2, 4,
                                                        5, 6],
                                                    itinerariomalla=itinerariosexto).exists()
                                                practicarechazada = False

                                                if not practica:
                                                    practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        estadosolicitud=3,
                                                        itinerariomalla=itinerariosexto).exists()

                                                if not practica or practicarechazada:
                                                    if not PracticasPreprofesionalesInscripcion.objects.filter(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            actividad__itinerariomalla=itinerariosexto).exists():
                                                        nuevapractica = PracticasPreprofesionalesInscripcion(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                            fechahasta=recordnuevo.materiaregular.nivel.periodo.fin,
                                                            numerohora=isextonuevo.horas_practicas,
                                                            nivelmalla=isextonuevo.nivel,
                                                            tiposolicitud=1,
                                                            estadosolicitud=2,
                                                            tipo=1,
                                                            itinerariomalla=isextonuevo,
                                                            supervisor=profesor,
                                                            tutorunemi=profesor,
                                                            fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                            tipoinstitucion=1,
                                                            sectoreconomico=6,
                                                            empresaempleadora_id=3,
                                                            culminada=True,
                                                            fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                            lugarpractica_id=2,
                                                            observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                        )
                                                        nuevapractica.save()

                                                ItinerariosActividadDetalleDistributivoCarrera.objects.filter(
                                                    status=True,
                                                    itinerario=itinerariosexto).update(
                                                    itinerario=isextonuevo)

                                            if equivalencia.asignaturamallasalto_id == 10953:
                                                itinerarioseptimo = ItinerariosMalla.objects.get(status=True,
                                                                                                 malla_id=mallaantigua_id,
                                                                                                 nivel_id=7)
                                                iseptimonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                             malla_id=mallanueva_id,
                                                                                             nivel_id=7)

                                                practica = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud__in=[
                                                        1, 2, 4,
                                                        5, 6],
                                                    itinerariomalla=itinerarioseptimo).exists()
                                                practicarechazada = False

                                                if not practica:
                                                    practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        estadosolicitud=3,
                                                        itinerariomalla=itinerarioseptimo).exists()

                                                if not practica or practicarechazada:
                                                    if not PracticasPreprofesionalesInscripcion.objects.filter(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            actividad__itinerariomalla=itinerarioseptimo).exists():
                                                        nuevapractica = PracticasPreprofesionalesInscripcion(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                            fechahasta=recordnuevo.materiaregular.nivel.periodo.fin,
                                                            numerohora=iseptimonuevo.horas_practicas,
                                                            nivelmalla=iseptimonuevo.nivel,
                                                            tiposolicitud=1,
                                                            estadosolicitud=2,
                                                            tipo=1,
                                                            itinerariomalla=iseptimonuevo,
                                                            supervisor=profesor,
                                                            tutorunemi=profesor,
                                                            fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                            tipoinstitucion=1,
                                                            sectoreconomico=6,
                                                            empresaempleadora_id=3,
                                                            culminada=True,
                                                            fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                            lugarpractica_id=2,
                                                            observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                        )
                                                        nuevapractica.save()

                                if equivalencia.asignaturamallasalto_id in [10959, 10964]:
                                    if not horasvinculacion:
                                        if recordnuevo.aprobada:
                                            totalhoras = inscripcion.numero_horas_proyectos_vinculacion()
                                            if totalhoras == 0:
                                                a = 0
                                                print('sin horas')
                                            if equivalencia.asignaturamallasalto_id == 10959 and inscripcion.numero_horas_proyectos_vinculacion() < 80:
                                                horasfalta = 80 - inscripcion.numero_horas_proyectos_vinculacion()
                                                vinculacion = ParticipantesMatrices(status=True,
                                                                                    matrizevidencia_id=2,
                                                                                    proyecto_id=601,
                                                                                    inscripcion=inscripcion,
                                                                                    horas=horasfalta,
                                                                                    registrohorasdesde=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                                                    registrohorashasta=recordnuevo.materiaregular.nivel.periodo.fin,
                                                                                    estado=1
                                                                                    )
                                                vinculacion.save()

                                            if equivalencia.asignaturamallasalto_id == 10964 and inscripcion.numero_horas_proyectos_vinculacion() < 160:
                                                horasfalta = 160 - inscripcion.numero_horas_proyectos_vinculacion()
                                                vinculacion = ParticipantesMatrices(status=True,
                                                                                    matrizevidencia_id=2,
                                                                                    proyecto_id=601,
                                                                                    inscripcion=inscripcion,
                                                                                    horas=horasfalta,
                                                                                    registrohorasdesde=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                                                    registrohorashasta=recordnuevo.materiaregular.nivel.periodo.fin,
                                                                                    estado=1
                                                                                    )
                                                vinculacion.save()

                                if not respaldo.exists():
                                    respaldorecord = RespaldoRecordAcademico(
                                        recordacademicooriginal=recordantiguo,
                                        recordacademiconuevo=recordnuevo
                                    )
                                    respaldorecord.save()
                                else:
                                    respaldorecord = respaldo[0]
                                    respaldorecord.recordacademiconuevo = recordnuevo
                                    respaldorecord.save()
                                print(u"Record actualizado %s" % recordnuevo)


                        else:
                            hojadestino.write(fila, 3, "NO ENCONTRO RECORD ANTIGUO %s" % equivalencia.asignaturamalla)
                            fila += 1

                    fila += 1

                    time.sleep(3)

            lin += 1

        libdestino.save(output_folder + libre_origen)
        print(output_folder + libre_origen)
        print("Proceso finalizado. . .")


    except Exception as ex:

        transaction.set_rollback(True)
        print('error: %s' % ex)
        hojadestino.write(fila, 3, str(ex))
        fila += 1

def homologacion_turismo_4():
    try:
        libre_origen = '/homologacion_turismo_2_4.xls'
        fuentecabecera = easyxf(
            'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
        fuentenormal = easyxf(
            'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')

        output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
        # liborigen = xlrd.open_workbook(output_folder + libre_origen)
        libdestino = xlwt.Workbook()
        hojadestino = libdestino.add_sheet('Sheet1')
        fil = 0
        columnas = [(u"CEDULA", 6000),
                    (u"APELLIDOS Y NOMBRES", 6000),
                    (u"OBSERVACIÓN", 6000)
                    ]
        for col_num in range(len(columnas)):
            hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
            hojadestino.col(col_num).width = columnas[col_num][1]
        fila = 1
        titulacion = 0

        lin = 0
        # miarchivo = openpyxl.load_workbook("CEDULA_COMUNICACION_final.xlsx")
        miarchivo = openpyxl.load_workbook("turismolinea.xlsx")
        # miarchivo = openpyxl.load_workbook(path_anexo)
        # miarchivo = openpyxl.load_workbook("prueba_2.xlsx")

        ws = miarchivo.get_sheet_by_name("cuarto")
        worksheet = ws
        c = 0
        cont = 0
        periodo_id = 224
        carrera_id = 134
        mallaantigua_id = 199
        mallanueva_id = 487
        for row in worksheet.iter_rows(min_row=0):
            if lin >= 0:
                currentValues, cadena = [], ''
                for cell in row:
                    cadena += str(cell.value) + ' '
                    currentValues.append(str(cell.value))
                identificacion = currentValues[0]

                if not identificacion:
                    break

                with transaction.atomic():
                    matricula = Matricula.objects.filter(status=True, inscripcion__carrera__id=carrera_id,
                                                         inscripcion__persona__cedula=identificacion).first()
                    cont += 1
                    matricula.pasoayuda = True
                    matricula.save()
                    print(u"%s - %s" % (matricula, cont))
                    inscripcion = matricula.inscripcion
                    hojadestino.write(fila, 0, matricula.inscripcion.persona.identificacion(), fuentenormal)
                    hojadestino.write(fila, 1, matricula.inscripcion.persona.nombre_completo(), fuentenormal)
                    hojadestino.write(fila, 2, matricula.nivel.periodo.nombre, fuentenormal)

                    itinerarios = ItinerariosMalla.objects.filter(status=True, malla_id=mallaantigua_id)

                    for itinerario in itinerarios:
                        itinerarionuevo = ItinerariosMalla.objects.filter(status=True, malla_id=mallanueva_id,
                                                                          nivel_id=itinerario.nivel_id)
                        # estadodo culmiado, en curso y pendiente
                        practicasencurso = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                               inscripcion=inscripcion,
                                                                                               estadosolicitud=2,
                                                                                               culminada=False,
                                                                                               itinerariomalla=itinerario)
                        for pcurso in practicasencurso:
                            pcurso.culminada = True
                            pcurso.save()

                        practicaspendientes = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                  inscripcion=inscripcion,
                                                                                                  estadosolicitud=4,
                                                                                                  culminada=False,
                                                                                                  itinerariomalla=itinerario)
                        for pendiente in practicaspendientes:
                            pendiente.estadosolicitud = 2
                            pendiente.culminada = True
                            pendiente.save()

                    practicaspp = haber_cumplido_horas_creditos_practicas_preprofesionales(inscripcion.id)
                    horasvinculacion = haber_cumplido_horas_creditos_vinculacion(inscripcion.id)

                    if InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                       malla_id=mallaantigua_id).exists():
                        imantigua = \
                            InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                            malla_id=mallaantigua_id)[
                                0]
                        imantigua.status = False
                        imantigua.save()
                        print(u"Desactiva antigua inscripcion -----------------------------")

                    if not InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                           malla_id=mallanueva_id).exists():
                        imnueva = InscripcionMalla(inscripcion=inscripcion, malla_id=mallanueva_id)
                        imnueva.save()
                        print(u"Crea nueva inscripcion -----------------------------")

                    equivalencias = TablaEquivalenciaAsignaturas.objects.filter(status=True,
                                                                                asignaturamalla__malla_id=mallaantigua_id).order_by(
                        'asignaturamallasalto__nivelmalla__orden')
                    for equivalencia in equivalencias:
                        print(u"nueva - %s" % equivalencia.asignaturamallasalto)
                        recordantiguo = inscripcion.recordacademico_set.filter(status=True,
                                                                               asignaturamalla_id=equivalencia.asignaturamalla_id).first()

                        if recordantiguo:
                            print(u"anterior - %s" % equivalencia.asignaturamalla)
                            print(u"Record antiguo: %s" % recordantiguo)
                            recordnuevo = None
                            recordantiguo.status = False
                            recordantiguo.save(update_asignaturamalla=False)

                            if equivalencia.asignaturamallasalto_id in [10929, 10947, 10959, 10953, 10964]:
                                observaciones = recordantiguo.observaciones + " Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15"
                                homologada = True



                            else:
                                observaciones = recordantiguo.observaciones + " Migración con base a RESOLUCIÓN CGA-SO-9-2023-NO15 Y RESOLUCIÓN OCS-SO-18-2023-NO14"
                                homologada = False
                            if not RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                  asignaturamalla=equivalencia.asignaturamallasalto).exists():

                                recordnuevo = RecordAcademico(inscripcion=inscripcion,
                                                              matriculas=recordantiguo.matriculas,
                                                              asignaturamalla=equivalencia.asignaturamallasalto,
                                                              asignatura=equivalencia.asignaturamallasalto.asignatura,
                                                              asignaturaold_id=recordantiguo.asignatura.id,
                                                              nota=recordantiguo.nota,
                                                              asistencia=recordantiguo.asistencia,
                                                              sinasistencia=recordantiguo.sinasistencia,
                                                              fecha=recordantiguo.fecha,
                                                              noaplica=recordantiguo.noaplica,
                                                              aprobada=recordantiguo.aprobada,
                                                              convalidacion=recordantiguo.convalidacion,
                                                              pendiente=recordantiguo.pendiente,
                                                              creditos=equivalencia.asignaturamallasalto.creditos,
                                                              horas=equivalencia.asignaturamallasalto.horas,
                                                              valida=recordantiguo.valida,
                                                              validapromedio=recordantiguo.validapromedio,
                                                              observaciones=observaciones,
                                                              homologada=homologada,
                                                              materiaregular=recordantiguo.materiaregular,
                                                              materiacurso=None,
                                                              completonota=recordantiguo.completonota,
                                                              completoasistencia=recordantiguo.completoasistencia,
                                                              fechainicio=recordantiguo.fechainicio,
                                                              fechafin=recordantiguo.fechafin,
                                                              suficiencia=recordantiguo.suficiencia,
                                                              asignaturamallahistorico_id=recordantiguo.asignaturamalla.id,
                                                              reverso=False)
                                recordnuevo.save()
                                print(u"Crea nuevo record %s" % recordnuevo)


                            elif RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                asignaturamalla=equivalencia.asignaturamallasalto):
                                recordnuevo = RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                             asignaturamalla=equivalencia.asignaturamallasalto)[
                                    0]
                                recordnuevo.matriculas = recordantiguo.matriculas
                                recordnuevo.asignaturamalla = equivalencia.asignaturamallasalto
                                recordnuevo.asignatura = equivalencia.asignaturamallasalto.asignatura
                                recordnuevo.asignaturaold = recordantiguo.asignaturamalla.asignatura
                                recordnuevo.nota = recordantiguo.nota
                                recordnuevo.asistencia = recordantiguo.asistencia
                                recordnuevo.sinasistencia = recordantiguo.sinasistencia
                                recordnuevo.fecha = recordantiguo.fecha
                                recordnuevo.noaplica = recordantiguo.noaplica
                                recordnuevo.aprobada = recordantiguo.aprobada
                                recordnuevo.convalidacion = recordantiguo.convalidacion
                                recordnuevo.pendiente = recordantiguo.pendiente
                                recordnuevo.creditos = equivalencia.asignaturamallasalto.creditos
                                recordnuevo.horas = equivalencia.asignaturamallasalto.horas
                                recordnuevo.valida = recordantiguo.valida
                                recordnuevo.validapromedio = recordantiguo.validapromedio
                                recordnuevo.observaciones = observaciones
                                recordnuevo.homologada = homologada
                                recordnuevo.materiaregular = recordantiguo.materiaregular
                                recordnuevo.materiacurso = None
                                recordnuevo.completonota = recordantiguo.completonota
                                recordnuevo.completoasistencia = recordantiguo.completoasistencia
                                recordnuevo.fechainicio = recordantiguo.fechainicio
                                recordnuevo.fechafin = recordantiguo.fechafin
                                recordnuevo.suficiencia = recordantiguo.suficiencia
                                recordnuevo.asignaturamallahistorico = recordantiguo.asignaturamalla
                                recordnuevo.reverso = False
                                recordnuevo.save()

                            if recordnuevo:
                                historicos = HistoricoRecordAcademico.objects.filter(status=True,
                                                                                     recordacademico=recordantiguo).update(
                                    recordacademico=recordnuevo,
                                    creditos=recordnuevo.creditos,
                                    horas=recordnuevo.horas,
                                    homologada=recordnuevo.homologada)
                                respaldo = RespaldoRecordAcademico.objects.filter(status=True,
                                                                                  recordacademicooriginal=recordantiguo)

                                if equivalencia.asignaturamallasalto_id in [10929, 10947, 10953]:
                                    if not practicaspp:
                                        if recordnuevo.aprobada:
                                            profesor = None
                                            if recordnuevo.materiaregular:
                                                profesor = recordnuevo.materiaregular.profesor_principal()
                                            elif recordnuevo.materiacurso:
                                                profesor = recordnuevo.materiaregular.profesor()
                                            if equivalencia.asignaturamallasalto_id == 10929:
                                                itinerarioquinto = ItinerariosMalla.objects.get(status=True,
                                                                                                malla_id=mallaantigua_id,
                                                                                                nivel_id=5)
                                                iquintonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                            malla_id=mallanueva_id,
                                                                                            nivel_id=5)

                                                practica = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud__in=[
                                                        1, 2, 4,
                                                        5, 6],
                                                    itinerariomalla=itinerarioquinto).exists()
                                                practicarechazada = False

                                                if not practica:
                                                    practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        estadosolicitud=3,
                                                        itinerariomalla=itinerarioquinto).exists()

                                                if not practica or practicarechazada:
                                                    if not PracticasPreprofesionalesInscripcion.objects.filter(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            actividad__itinerariomalla=itinerarioquinto).exists():
                                                        nuevapractica = PracticasPreprofesionalesInscripcion(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            numerohora=iquintonuevo.horas_practicas,
                                                            nivelmalla=iquintonuevo.nivel,
                                                            tiposolicitud=1,
                                                            estadosolicitud=2,
                                                            tipo=1,
                                                            itinerariomalla=iquintonuevo,
                                                            supervisor=profesor,
                                                            tutorunemi=profesor,
                                                            fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            tipoinstitucion=1,
                                                            sectoreconomico=6,
                                                            empresaempleadora_id=3,
                                                            culminada=True,
                                                            fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            lugarpractica_id=2,
                                                            observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                        )
                                                        nuevapractica.save()

                                                ItinerariosActividadDetalleDistributivoCarrera.objects.filter(
                                                    status=True,
                                                    itinerario=itinerarioquinto).update(
                                                    itinerario=iquintonuevo)

                                            if equivalencia.asignaturamallasalto_id == 10947:
                                                itinerariosexto = ItinerariosMalla.objects.get(status=True,
                                                                                               malla_id=mallaantigua_id,
                                                                                               nivel_id=6)
                                                isextonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallanueva_id,
                                                                                           nivel_id=6)

                                                practica = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud__in=[
                                                        1, 2, 4,
                                                        5, 6],
                                                    itinerariomalla=itinerariosexto).exists()
                                                practicarechazada = False

                                                if not practica:
                                                    practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        estadosolicitud=3,
                                                        itinerariomalla=itinerariosexto).exists()

                                                if not practica or practicarechazada:
                                                    if not PracticasPreprofesionalesInscripcion.objects.filter(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            actividad__itinerariomalla=itinerariosexto).exists():
                                                        nuevapractica = PracticasPreprofesionalesInscripcion(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                            fechahasta=recordnuevo.materiaregular.nivel.periodo.fin,
                                                            numerohora=isextonuevo.horas_practicas,
                                                            nivelmalla=isextonuevo.nivel,
                                                            tiposolicitud=1,
                                                            estadosolicitud=2,
                                                            tipo=1,
                                                            itinerariomalla=isextonuevo,
                                                            supervisor=profesor,
                                                            tutorunemi=profesor,
                                                            fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                            tipoinstitucion=1,
                                                            sectoreconomico=6,
                                                            empresaempleadora_id=3,
                                                            culminada=True,
                                                            fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                            lugarpractica_id=2,
                                                            observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                        )
                                                        nuevapractica.save()

                                                ItinerariosActividadDetalleDistributivoCarrera.objects.filter(
                                                    status=True,
                                                    itinerario=itinerariosexto).update(
                                                    itinerario=isextonuevo)

                                            if equivalencia.asignaturamallasalto_id == 10953:
                                                itinerarioseptimo = ItinerariosMalla.objects.get(status=True,
                                                                                                 malla_id=mallaantigua_id,
                                                                                                 nivel_id=7)
                                                iseptimonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                             malla_id=mallanueva_id,
                                                                                             nivel_id=7)

                                                practica = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud__in=[
                                                        1, 2, 4,
                                                        5, 6],
                                                    itinerariomalla=itinerarioseptimo).exists()
                                                practicarechazada = False

                                                if not practica:
                                                    practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        estadosolicitud=3,
                                                        itinerariomalla=itinerarioseptimo).exists()

                                                if not practica or practicarechazada:
                                                    if not PracticasPreprofesionalesInscripcion.objects.filter(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            actividad__itinerariomalla=itinerarioseptimo).exists():
                                                        nuevapractica = PracticasPreprofesionalesInscripcion(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                            fechahasta=recordnuevo.materiaregular.nivel.periodo.fin,
                                                            numerohora=iseptimonuevo.horas_practicas,
                                                            nivelmalla=iseptimonuevo.nivel,
                                                            tiposolicitud=1,
                                                            estadosolicitud=2,
                                                            tipo=1,
                                                            itinerariomalla=iseptimonuevo,
                                                            supervisor=profesor,
                                                            tutorunemi=profesor,
                                                            fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                            tipoinstitucion=1,
                                                            sectoreconomico=6,
                                                            empresaempleadora_id=3,
                                                            culminada=True,
                                                            fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                            lugarpractica_id=2,
                                                            observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                        )
                                                        nuevapractica.save()

                                if equivalencia.asignaturamallasalto_id in [10959, 10964]:
                                    if not horasvinculacion:
                                        if recordnuevo.aprobada:
                                            totalhoras = inscripcion.numero_horas_proyectos_vinculacion()
                                            if totalhoras == 0:
                                                a = 0
                                                print('sin horas')
                                            if equivalencia.asignaturamallasalto_id == 10959 and inscripcion.numero_horas_proyectos_vinculacion() < 80:
                                                horasfalta = 80 - inscripcion.numero_horas_proyectos_vinculacion()
                                                vinculacion = ParticipantesMatrices(status=True,
                                                                                    matrizevidencia_id=2,
                                                                                    proyecto_id=601,
                                                                                    inscripcion=inscripcion,
                                                                                    horas=horasfalta,
                                                                                    registrohorasdesde=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                                                    registrohorashasta=recordnuevo.materiaregular.nivel.periodo.fin,
                                                                                    estado=1
                                                                                    )
                                                vinculacion.save()

                                            if equivalencia.asignaturamallasalto_id == 10964 and inscripcion.numero_horas_proyectos_vinculacion() < 160:
                                                horasfalta = 160 - inscripcion.numero_horas_proyectos_vinculacion()
                                                vinculacion = ParticipantesMatrices(status=True,
                                                                                    matrizevidencia_id=2,
                                                                                    proyecto_id=601,
                                                                                    inscripcion=inscripcion,
                                                                                    horas=horasfalta,
                                                                                    registrohorasdesde=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                                                    registrohorashasta=recordnuevo.materiaregular.nivel.periodo.fin,
                                                                                    estado=1
                                                                                    )
                                                vinculacion.save()

                                if not respaldo.exists():
                                    respaldorecord = RespaldoRecordAcademico(
                                        recordacademicooriginal=recordantiguo,
                                        recordacademiconuevo=recordnuevo
                                    )
                                    respaldorecord.save()
                                else:
                                    respaldorecord = respaldo[0]
                                    respaldorecord.recordacademiconuevo = recordnuevo
                                    respaldorecord.save()
                                print(u"Record actualizado %s" % recordnuevo)


                        else:
                            hojadestino.write(fila, 3, "NO ENCONTRO RECORD ANTIGUO %s" % equivalencia.asignaturamalla)
                            fila += 1

                    fila += 1

                    time.sleep(3)

            lin += 1

        libdestino.save(output_folder + libre_origen)
        print(output_folder + libre_origen)
        print("Proceso finalizado. . .")


    except Exception as ex:

        transaction.set_rollback(True)
        print('error: %s' % ex)
        hojadestino.write(fila, 3, str(ex))
        fila += 1

def homologacion_turismo_5():
    try:
        libre_origen = '/homologacion_turismo_2_5.xls'
        fuentecabecera = easyxf(
            'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
        fuentenormal = easyxf(
            'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')

        output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
        # liborigen = xlrd.open_workbook(output_folder + libre_origen)
        libdestino = xlwt.Workbook()
        hojadestino = libdestino.add_sheet('Sheet1')
        fil = 0
        columnas = [(u"CEDULA", 6000),
                    (u"APELLIDOS Y NOMBRES", 6000),
                    (u"OBSERVACIÓN", 6000)
                    ]
        for col_num in range(len(columnas)):
            hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
            hojadestino.col(col_num).width = columnas[col_num][1]
        fila = 1
        titulacion = 0

        lin = 0
        # miarchivo = openpyxl.load_workbook("CEDULA_COMUNICACION_final.xlsx")
        miarchivo = openpyxl.load_workbook("turismolinea.xlsx")
        # miarchivo = openpyxl.load_workbook(path_anexo)
        # miarchivo = openpyxl.load_workbook("prueba_2.xlsx")

        ws = miarchivo.get_sheet_by_name("quinto")
        worksheet = ws
        c = 0
        cont = 0
        periodo_id = 224
        carrera_id = 134
        mallaantigua_id = 199
        mallanueva_id = 487
        for row in worksheet.iter_rows(min_row=0):
            if lin >= 0:
                currentValues, cadena = [], ''
                for cell in row:
                    cadena += str(cell.value) + ' '
                    currentValues.append(str(cell.value))
                identificacion = currentValues[0]

                if not identificacion:
                    break

                with transaction.atomic():
                    matricula = Matricula.objects.filter(status=True, inscripcion__carrera__id=carrera_id,
                                                         inscripcion__persona__cedula=identificacion).first()
                    cont += 1
                    matricula.pasoayuda = True
                    matricula.save()
                    print(u"%s - %s" % (matricula, cont))
                    inscripcion = matricula.inscripcion
                    hojadestino.write(fila, 0, matricula.inscripcion.persona.identificacion(), fuentenormal)
                    hojadestino.write(fila, 1, matricula.inscripcion.persona.nombre_completo(), fuentenormal)
                    hojadestino.write(fila, 2, matricula.nivel.periodo.nombre, fuentenormal)

                    itinerarios = ItinerariosMalla.objects.filter(status=True, malla_id=mallaantigua_id)

                    for itinerario in itinerarios:
                        itinerarionuevo = ItinerariosMalla.objects.filter(status=True, malla_id=mallanueva_id,
                                                                          nivel_id=itinerario.nivel_id)
                        # estadodo culmiado, en curso y pendiente
                        practicasencurso = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                               inscripcion=inscripcion,
                                                                                               estadosolicitud=2,
                                                                                               culminada=False,
                                                                                               itinerariomalla=itinerario)
                        for pcurso in practicasencurso:
                            pcurso.culminada = True
                            pcurso.save()

                        practicaspendientes = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                  inscripcion=inscripcion,
                                                                                                  estadosolicitud=4,
                                                                                                  culminada=False,
                                                                                                  itinerariomalla=itinerario)
                        for pendiente in practicaspendientes:
                            pendiente.estadosolicitud = 2
                            pendiente.culminada = True
                            pendiente.save()

                    practicaspp = haber_cumplido_horas_creditos_practicas_preprofesionales(inscripcion.id)
                    horasvinculacion = haber_cumplido_horas_creditos_vinculacion(inscripcion.id)

                    if InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                       malla_id=mallaantigua_id).exists():
                        imantigua = \
                            InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                            malla_id=mallaantigua_id)[
                                0]
                        imantigua.status = False
                        imantigua.save()
                        print(u"Desactiva antigua inscripcion -----------------------------")

                    if not InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                           malla_id=mallanueva_id).exists():
                        imnueva = InscripcionMalla(inscripcion=inscripcion, malla_id=mallanueva_id)
                        imnueva.save()
                        print(u"Crea nueva inscripcion -----------------------------")

                    equivalencias = TablaEquivalenciaAsignaturas.objects.filter(status=True,
                                                                                asignaturamalla__malla_id=mallaantigua_id).order_by(
                        'asignaturamallasalto__nivelmalla__orden')
                    for equivalencia in equivalencias:
                        print(u"nueva - %s" % equivalencia.asignaturamallasalto)
                        recordantiguo = inscripcion.recordacademico_set.filter(status=True,
                                                                               asignaturamalla_id=equivalencia.asignaturamalla_id).first()

                        if recordantiguo:
                            print(u"anterior - %s" % equivalencia.asignaturamalla)
                            print(u"Record antiguo: %s" % recordantiguo)
                            recordnuevo = None
                            recordantiguo.status = False
                            recordantiguo.save(update_asignaturamalla=False)

                            if equivalencia.asignaturamallasalto_id in [10929, 10947, 10959, 10953, 10964]:
                                observaciones = recordantiguo.observaciones + " Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15"
                                homologada = True



                            else:
                                observaciones = recordantiguo.observaciones + " Migración con base a RESOLUCIÓN CGA-SO-9-2023-NO15 Y RESOLUCIÓN OCS-SO-18-2023-NO14"
                                homologada = False
                            if not RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                  asignaturamalla=equivalencia.asignaturamallasalto).exists():

                                recordnuevo = RecordAcademico(inscripcion=inscripcion,
                                                              matriculas=recordantiguo.matriculas,
                                                              asignaturamalla=equivalencia.asignaturamallasalto,
                                                              asignatura=equivalencia.asignaturamallasalto.asignatura,
                                                              asignaturaold_id=recordantiguo.asignatura.id,
                                                              nota=recordantiguo.nota,
                                                              asistencia=recordantiguo.asistencia,
                                                              sinasistencia=recordantiguo.sinasistencia,
                                                              fecha=recordantiguo.fecha,
                                                              noaplica=recordantiguo.noaplica,
                                                              aprobada=recordantiguo.aprobada,
                                                              convalidacion=recordantiguo.convalidacion,
                                                              pendiente=recordantiguo.pendiente,
                                                              creditos=equivalencia.asignaturamallasalto.creditos,
                                                              horas=equivalencia.asignaturamallasalto.horas,
                                                              valida=recordantiguo.valida,
                                                              validapromedio=recordantiguo.validapromedio,
                                                              observaciones=observaciones,
                                                              homologada=homologada,
                                                              materiaregular=recordantiguo.materiaregular,
                                                              materiacurso=None,
                                                              completonota=recordantiguo.completonota,
                                                              completoasistencia=recordantiguo.completoasistencia,
                                                              fechainicio=recordantiguo.fechainicio,
                                                              fechafin=recordantiguo.fechafin,
                                                              suficiencia=recordantiguo.suficiencia,
                                                              asignaturamallahistorico_id=recordantiguo.asignaturamalla.id,
                                                              reverso=False)
                                recordnuevo.save()
                                print(u"Crea nuevo record %s" % recordnuevo)


                            elif RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                asignaturamalla=equivalencia.asignaturamallasalto):
                                recordnuevo = RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                             asignaturamalla=equivalencia.asignaturamallasalto)[
                                    0]
                                recordnuevo.matriculas = recordantiguo.matriculas
                                recordnuevo.asignaturamalla = equivalencia.asignaturamallasalto
                                recordnuevo.asignatura = equivalencia.asignaturamallasalto.asignatura
                                recordnuevo.asignaturaold = recordantiguo.asignaturamalla.asignatura
                                recordnuevo.nota = recordantiguo.nota
                                recordnuevo.asistencia = recordantiguo.asistencia
                                recordnuevo.sinasistencia = recordantiguo.sinasistencia
                                recordnuevo.fecha = recordantiguo.fecha
                                recordnuevo.noaplica = recordantiguo.noaplica
                                recordnuevo.aprobada = recordantiguo.aprobada
                                recordnuevo.convalidacion = recordantiguo.convalidacion
                                recordnuevo.pendiente = recordantiguo.pendiente
                                recordnuevo.creditos = equivalencia.asignaturamallasalto.creditos
                                recordnuevo.horas = equivalencia.asignaturamallasalto.horas
                                recordnuevo.valida = recordantiguo.valida
                                recordnuevo.validapromedio = recordantiguo.validapromedio
                                recordnuevo.observaciones = observaciones
                                recordnuevo.homologada = homologada
                                recordnuevo.materiaregular = recordantiguo.materiaregular
                                recordnuevo.materiacurso = None
                                recordnuevo.completonota = recordantiguo.completonota
                                recordnuevo.completoasistencia = recordantiguo.completoasistencia
                                recordnuevo.fechainicio = recordantiguo.fechainicio
                                recordnuevo.fechafin = recordantiguo.fechafin
                                recordnuevo.suficiencia = recordantiguo.suficiencia
                                recordnuevo.asignaturamallahistorico = recordantiguo.asignaturamalla
                                recordnuevo.reverso = False
                                recordnuevo.save()

                            if recordnuevo:
                                historicos = HistoricoRecordAcademico.objects.filter(status=True,
                                                                                     recordacademico=recordantiguo).update(
                                    recordacademico=recordnuevo,
                                    creditos=recordnuevo.creditos,
                                    horas=recordnuevo.horas,
                                    homologada=recordnuevo.homologada)
                                respaldo = RespaldoRecordAcademico.objects.filter(status=True,
                                                                                  recordacademicooriginal=recordantiguo)

                                if equivalencia.asignaturamallasalto_id in [10929, 10947, 10953]:
                                    if not practicaspp:
                                        if recordnuevo.aprobada:
                                            profesor = None
                                            if recordnuevo.materiaregular:
                                                profesor = recordnuevo.materiaregular.profesor_principal()
                                            elif recordnuevo.materiacurso:
                                                profesor = recordnuevo.materiaregular.profesor()
                                            if equivalencia.asignaturamallasalto_id == 10929:
                                                itinerarioquinto = ItinerariosMalla.objects.get(status=True,
                                                                                                malla_id=mallaantigua_id,
                                                                                                nivel_id=5)
                                                iquintonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                            malla_id=mallanueva_id,
                                                                                            nivel_id=5)

                                                practica = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud__in=[
                                                        1, 2, 4,
                                                        5, 6],
                                                    itinerariomalla=itinerarioquinto).exists()
                                                practicarechazada = False

                                                if not practica:
                                                    practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        estadosolicitud=3,
                                                        itinerariomalla=itinerarioquinto).exists()

                                                if not practica or practicarechazada:
                                                    if not PracticasPreprofesionalesInscripcion.objects.filter(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            actividad__itinerariomalla=itinerarioquinto).exists():
                                                        nuevapractica = PracticasPreprofesionalesInscripcion(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            numerohora=iquintonuevo.horas_practicas,
                                                            nivelmalla=iquintonuevo.nivel,
                                                            tiposolicitud=1,
                                                            estadosolicitud=2,
                                                            tipo=1,
                                                            itinerariomalla=iquintonuevo,
                                                            supervisor=profesor,
                                                            tutorunemi=profesor,
                                                            fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            tipoinstitucion=1,
                                                            sectoreconomico=6,
                                                            empresaempleadora_id=3,
                                                            culminada=True,
                                                            fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            lugarpractica_id=2,
                                                            observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                        )
                                                        nuevapractica.save()

                                                ItinerariosActividadDetalleDistributivoCarrera.objects.filter(
                                                    status=True,
                                                    itinerario=itinerarioquinto).update(
                                                    itinerario=iquintonuevo)

                                            if equivalencia.asignaturamallasalto_id == 10947:
                                                itinerariosexto = ItinerariosMalla.objects.get(status=True,
                                                                                               malla_id=mallaantigua_id,
                                                                                               nivel_id=6)
                                                isextonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallanueva_id,
                                                                                           nivel_id=6)

                                                practica = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud__in=[
                                                        1, 2, 4,
                                                        5, 6],
                                                    itinerariomalla=itinerariosexto).exists()
                                                practicarechazada = False

                                                if not practica:
                                                    practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        estadosolicitud=3,
                                                        itinerariomalla=itinerariosexto).exists()

                                                if not practica or practicarechazada:
                                                    if not PracticasPreprofesionalesInscripcion.objects.filter(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            actividad__itinerariomalla=itinerariosexto).exists():
                                                        nuevapractica = PracticasPreprofesionalesInscripcion(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                            fechahasta=recordnuevo.materiaregular.nivel.periodo.fin,
                                                            numerohora=isextonuevo.horas_practicas,
                                                            nivelmalla=isextonuevo.nivel,
                                                            tiposolicitud=1,
                                                            estadosolicitud=2,
                                                            tipo=1,
                                                            itinerariomalla=isextonuevo,
                                                            supervisor=profesor,
                                                            tutorunemi=profesor,
                                                            fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                            tipoinstitucion=1,
                                                            sectoreconomico=6,
                                                            empresaempleadora_id=3,
                                                            culminada=True,
                                                            fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                            lugarpractica_id=2,
                                                            observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                        )
                                                        nuevapractica.save()

                                                ItinerariosActividadDetalleDistributivoCarrera.objects.filter(
                                                    status=True,
                                                    itinerario=itinerariosexto).update(
                                                    itinerario=isextonuevo)

                                            if equivalencia.asignaturamallasalto_id == 10953:
                                                itinerarioseptimo = ItinerariosMalla.objects.get(status=True,
                                                                                                 malla_id=mallaantigua_id,
                                                                                                 nivel_id=7)
                                                iseptimonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                             malla_id=mallanueva_id,
                                                                                             nivel_id=7)

                                                practica = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud__in=[
                                                        1, 2, 4,
                                                        5, 6],
                                                    itinerariomalla=itinerarioseptimo).exists()
                                                practicarechazada = False

                                                if not practica:
                                                    practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        estadosolicitud=3,
                                                        itinerariomalla=itinerarioseptimo).exists()

                                                if not practica or practicarechazada:
                                                    if not PracticasPreprofesionalesInscripcion.objects.filter(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            actividad__itinerariomalla=itinerarioseptimo).exists():
                                                        nuevapractica = PracticasPreprofesionalesInscripcion(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                            fechahasta=recordnuevo.materiaregular.nivel.periodo.fin,
                                                            numerohora=iseptimonuevo.horas_practicas,
                                                            nivelmalla=iseptimonuevo.nivel,
                                                            tiposolicitud=1,
                                                            estadosolicitud=2,
                                                            tipo=1,
                                                            itinerariomalla=iseptimonuevo,
                                                            supervisor=profesor,
                                                            tutorunemi=profesor,
                                                            fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                            tipoinstitucion=1,
                                                            sectoreconomico=6,
                                                            empresaempleadora_id=3,
                                                            culminada=True,
                                                            fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                            lugarpractica_id=2,
                                                            observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                        )
                                                        nuevapractica.save()

                                if equivalencia.asignaturamallasalto_id in [10959, 10964]:
                                    if not horasvinculacion:
                                        if recordnuevo.aprobada:
                                            totalhoras = inscripcion.numero_horas_proyectos_vinculacion()
                                            if totalhoras == 0:
                                                a = 0
                                                print('sin horas')
                                            if equivalencia.asignaturamallasalto_id == 10959 and inscripcion.numero_horas_proyectos_vinculacion() < 80:
                                                horasfalta = 80 - inscripcion.numero_horas_proyectos_vinculacion()
                                                vinculacion = ParticipantesMatrices(status=True,
                                                                                    matrizevidencia_id=2,
                                                                                    proyecto_id=601,
                                                                                    inscripcion=inscripcion,
                                                                                    horas=horasfalta,
                                                                                    registrohorasdesde=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                                                    registrohorashasta=recordnuevo.materiaregular.nivel.periodo.fin,
                                                                                    estado=1
                                                                                    )
                                                vinculacion.save()

                                            if equivalencia.asignaturamallasalto_id == 10964 and inscripcion.numero_horas_proyectos_vinculacion() < 160:
                                                horasfalta = 160 - inscripcion.numero_horas_proyectos_vinculacion()
                                                vinculacion = ParticipantesMatrices(status=True,
                                                                                    matrizevidencia_id=2,
                                                                                    proyecto_id=601,
                                                                                    inscripcion=inscripcion,
                                                                                    horas=horasfalta,
                                                                                    registrohorasdesde=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                                                    registrohorashasta=recordnuevo.materiaregular.nivel.periodo.fin,
                                                                                    estado=1
                                                                                    )
                                                vinculacion.save()

                                if not respaldo.exists():
                                    respaldorecord = RespaldoRecordAcademico(
                                        recordacademicooriginal=recordantiguo,
                                        recordacademiconuevo=recordnuevo
                                    )
                                    respaldorecord.save()
                                else:
                                    respaldorecord = respaldo[0]
                                    respaldorecord.recordacademiconuevo = recordnuevo
                                    respaldorecord.save()
                                print(u"Record actualizado %s" % recordnuevo)


                        else:
                            hojadestino.write(fila, 3, "NO ENCONTRO RECORD ANTIGUO %s" % equivalencia.asignaturamalla)
                            fila += 1

                    fila += 1

                    time.sleep(3)

            lin += 1

        libdestino.save(output_folder + libre_origen)
        print(output_folder + libre_origen)
        print("Proceso finalizado. . .")


    except Exception as ex:

        transaction.set_rollback(True)
        print('error: %s' % ex)
        hojadestino.write(fila, 3, str(ex))
        fila += 1

def homologacion_turismo_6():
    try:
        libre_origen = '/homologacion_turismo_2_6.xls'
        fuentecabecera = easyxf(
            'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
        fuentenormal = easyxf(
            'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')

        output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
        # liborigen = xlrd.open_workbook(output_folder + libre_origen)
        libdestino = xlwt.Workbook()
        hojadestino = libdestino.add_sheet('Sheet1')
        fil = 0
        columnas = [(u"CEDULA", 6000),
                    (u"APELLIDOS Y NOMBRES", 6000),
                    (u"OBSERVACIÓN", 6000)
                    ]
        for col_num in range(len(columnas)):
            hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
            hojadestino.col(col_num).width = columnas[col_num][1]
        fila = 1
        titulacion = 0

        lin = 0
        # miarchivo = openpyxl.load_workbook("CEDULA_COMUNICACION_final.xlsx")
        miarchivo = openpyxl.load_workbook("turismolinea.xlsx")
        # miarchivo = openpyxl.load_workbook(path_anexo)
        # miarchivo = openpyxl.load_workbook("prueba_2.xlsx")

        ws = miarchivo.get_sheet_by_name("sexto")
        worksheet = ws
        c = 0
        cont = 0
        periodo_id = 224
        carrera_id = 134
        mallaantigua_id = 199
        mallanueva_id = 487
        for row in worksheet.iter_rows(min_row=0):
            if lin >= 0:
                currentValues, cadena = [], ''
                for cell in row:
                    cadena += str(cell.value) + ' '
                    currentValues.append(str(cell.value))
                identificacion = currentValues[0]

                if not identificacion:
                    break

                with transaction.atomic():
                    matricula = Matricula.objects.filter(status=True, inscripcion__carrera__id=carrera_id,
                                                         inscripcion__persona__cedula=identificacion).first()
                    cont += 1
                    matricula.pasoayuda = True
                    matricula.save()
                    print(u"%s - %s" % (matricula, cont))
                    inscripcion = matricula.inscripcion
                    hojadestino.write(fila, 0, matricula.inscripcion.persona.identificacion(), fuentenormal)
                    hojadestino.write(fila, 1, matricula.inscripcion.persona.nombre_completo(), fuentenormal)
                    hojadestino.write(fila, 2, matricula.nivel.periodo.nombre, fuentenormal)

                    itinerarios = ItinerariosMalla.objects.filter(status=True, malla_id=mallaantigua_id)

                    for itinerario in itinerarios:
                        itinerarionuevo = ItinerariosMalla.objects.filter(status=True, malla_id=mallanueva_id,
                                                                          nivel_id=itinerario.nivel_id)
                        # estadodo culmiado, en curso y pendiente
                        practicasencurso = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                               inscripcion=inscripcion,
                                                                                               estadosolicitud=2,
                                                                                               culminada=False,
                                                                                               itinerariomalla=itinerario)
                        for pcurso in practicasencurso:
                            pcurso.culminada = True
                            pcurso.save()

                        practicaspendientes = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                  inscripcion=inscripcion,
                                                                                                  estadosolicitud=4,
                                                                                                  culminada=False,
                                                                                                  itinerariomalla=itinerario)
                        for pendiente in practicaspendientes:
                            pendiente.estadosolicitud = 2
                            pendiente.culminada = True
                            pendiente.save()

                    practicaspp = haber_cumplido_horas_creditos_practicas_preprofesionales(inscripcion.id)
                    horasvinculacion = haber_cumplido_horas_creditos_vinculacion(inscripcion.id)

                    if InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                       malla_id=mallaantigua_id).exists():
                        imantigua = \
                            InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                            malla_id=mallaantigua_id)[
                                0]
                        imantigua.status = False
                        imantigua.save()
                        print(u"Desactiva antigua inscripcion -----------------------------")

                    if not InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                           malla_id=mallanueva_id).exists():
                        imnueva = InscripcionMalla(inscripcion=inscripcion, malla_id=mallanueva_id)
                        imnueva.save()
                        print(u"Crea nueva inscripcion -----------------------------")

                    equivalencias = TablaEquivalenciaAsignaturas.objects.filter(status=True,
                                                                                asignaturamalla__malla_id=mallaantigua_id).order_by(
                        'asignaturamallasalto__nivelmalla__orden')
                    for equivalencia in equivalencias:
                        print(u"nueva - %s" % equivalencia.asignaturamallasalto)
                        recordantiguo = inscripcion.recordacademico_set.filter(status=True,
                                                                               asignaturamalla_id=equivalencia.asignaturamalla_id).first()

                        if recordantiguo:
                            print(u"anterior - %s" % equivalencia.asignaturamalla)
                            print(u"Record antiguo: %s" % recordantiguo)
                            recordnuevo = None
                            recordantiguo.status = False
                            recordantiguo.save(update_asignaturamalla=False)

                            if equivalencia.asignaturamallasalto_id in [10929, 10947, 10959, 10953, 10964]:
                                observaciones = recordantiguo.observaciones + " Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15"
                                homologada = True



                            else:
                                observaciones = recordantiguo.observaciones + " Migración con base a RESOLUCIÓN CGA-SO-9-2023-NO15 Y RESOLUCIÓN OCS-SO-18-2023-NO14"
                                homologada = False
                            if not RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                  asignaturamalla=equivalencia.asignaturamallasalto).exists():

                                recordnuevo = RecordAcademico(inscripcion=inscripcion,
                                                              matriculas=recordantiguo.matriculas,
                                                              asignaturamalla=equivalencia.asignaturamallasalto,
                                                              asignatura=equivalencia.asignaturamallasalto.asignatura,
                                                              asignaturaold_id=recordantiguo.asignatura.id,
                                                              nota=recordantiguo.nota,
                                                              asistencia=recordantiguo.asistencia,
                                                              sinasistencia=recordantiguo.sinasistencia,
                                                              fecha=recordantiguo.fecha,
                                                              noaplica=recordantiguo.noaplica,
                                                              aprobada=recordantiguo.aprobada,
                                                              convalidacion=recordantiguo.convalidacion,
                                                              pendiente=recordantiguo.pendiente,
                                                              creditos=equivalencia.asignaturamallasalto.creditos,
                                                              horas=equivalencia.asignaturamallasalto.horas,
                                                              valida=recordantiguo.valida,
                                                              validapromedio=recordantiguo.validapromedio,
                                                              observaciones=observaciones,
                                                              homologada=homologada,
                                                              materiaregular=recordantiguo.materiaregular,
                                                              materiacurso=None,
                                                              completonota=recordantiguo.completonota,
                                                              completoasistencia=recordantiguo.completoasistencia,
                                                              fechainicio=recordantiguo.fechainicio,
                                                              fechafin=recordantiguo.fechafin,
                                                              suficiencia=recordantiguo.suficiencia,
                                                              asignaturamallahistorico_id=recordantiguo.asignaturamalla.id,
                                                              reverso=False)
                                recordnuevo.save()
                                print(u"Crea nuevo record %s" % recordnuevo)


                            elif RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                asignaturamalla=equivalencia.asignaturamallasalto):
                                recordnuevo = RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                             asignaturamalla=equivalencia.asignaturamallasalto)[
                                    0]
                                recordnuevo.matriculas = recordantiguo.matriculas
                                recordnuevo.asignaturamalla = equivalencia.asignaturamallasalto
                                recordnuevo.asignatura = equivalencia.asignaturamallasalto.asignatura
                                recordnuevo.asignaturaold = recordantiguo.asignaturamalla.asignatura
                                recordnuevo.nota = recordantiguo.nota
                                recordnuevo.asistencia = recordantiguo.asistencia
                                recordnuevo.sinasistencia = recordantiguo.sinasistencia
                                recordnuevo.fecha = recordantiguo.fecha
                                recordnuevo.noaplica = recordantiguo.noaplica
                                recordnuevo.aprobada = recordantiguo.aprobada
                                recordnuevo.convalidacion = recordantiguo.convalidacion
                                recordnuevo.pendiente = recordantiguo.pendiente
                                recordnuevo.creditos = equivalencia.asignaturamallasalto.creditos
                                recordnuevo.horas = equivalencia.asignaturamallasalto.horas
                                recordnuevo.valida = recordantiguo.valida
                                recordnuevo.validapromedio = recordantiguo.validapromedio
                                recordnuevo.observaciones = observaciones
                                recordnuevo.homologada = homologada
                                recordnuevo.materiaregular = recordantiguo.materiaregular
                                recordnuevo.materiacurso = None
                                recordnuevo.completonota = recordantiguo.completonota
                                recordnuevo.completoasistencia = recordantiguo.completoasistencia
                                recordnuevo.fechainicio = recordantiguo.fechainicio
                                recordnuevo.fechafin = recordantiguo.fechafin
                                recordnuevo.suficiencia = recordantiguo.suficiencia
                                recordnuevo.asignaturamallahistorico = recordantiguo.asignaturamalla
                                recordnuevo.reverso = False
                                recordnuevo.save()

                            if recordnuevo:
                                historicos = HistoricoRecordAcademico.objects.filter(status=True,
                                                                                     recordacademico=recordantiguo).update(
                                    recordacademico=recordnuevo,
                                    creditos=recordnuevo.creditos,
                                    horas=recordnuevo.horas,
                                    homologada=recordnuevo.homologada)
                                respaldo = RespaldoRecordAcademico.objects.filter(status=True,
                                                                                  recordacademicooriginal=recordantiguo)

                                if equivalencia.asignaturamallasalto_id in [10929, 10947, 10953]:
                                    if not practicaspp:
                                        if recordnuevo.aprobada:
                                            profesor = None
                                            if recordnuevo.materiaregular:
                                                profesor = recordnuevo.materiaregular.profesor_principal()
                                            elif recordnuevo.materiacurso:
                                                profesor = recordnuevo.materiaregular.profesor()
                                            if equivalencia.asignaturamallasalto_id == 10929:
                                                itinerarioquinto = ItinerariosMalla.objects.get(status=True,
                                                                                                malla_id=mallaantigua_id,
                                                                                                nivel_id=5)
                                                iquintonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                            malla_id=mallanueva_id,
                                                                                            nivel_id=5)

                                                practica = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud__in=[
                                                        1, 2, 4,
                                                        5, 6],
                                                    itinerariomalla=itinerarioquinto).exists()
                                                practicarechazada = False

                                                if not practica:
                                                    practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        estadosolicitud=3,
                                                        itinerariomalla=itinerarioquinto).exists()

                                                if not practica or practicarechazada:
                                                    if not PracticasPreprofesionalesInscripcion.objects.filter(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            actividad__itinerariomalla=itinerarioquinto).exists():
                                                        nuevapractica = PracticasPreprofesionalesInscripcion(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            numerohora=iquintonuevo.horas_practicas,
                                                            nivelmalla=iquintonuevo.nivel,
                                                            tiposolicitud=1,
                                                            estadosolicitud=2,
                                                            tipo=1,
                                                            itinerariomalla=iquintonuevo,
                                                            supervisor=profesor,
                                                            tutorunemi=profesor,
                                                            fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            tipoinstitucion=1,
                                                            sectoreconomico=6,
                                                            empresaempleadora_id=3,
                                                            culminada=True,
                                                            fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            lugarpractica_id=2,
                                                            observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                        )
                                                        nuevapractica.save()

                                                ItinerariosActividadDetalleDistributivoCarrera.objects.filter(
                                                    status=True,
                                                    itinerario=itinerarioquinto).update(
                                                    itinerario=iquintonuevo)

                                            if equivalencia.asignaturamallasalto_id == 10947:
                                                itinerariosexto = ItinerariosMalla.objects.get(status=True,
                                                                                               malla_id=mallaantigua_id,
                                                                                               nivel_id=6)
                                                isextonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallanueva_id,
                                                                                           nivel_id=6)

                                                practica = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud__in=[
                                                        1, 2, 4,
                                                        5, 6],
                                                    itinerariomalla=itinerariosexto).exists()
                                                practicarechazada = False

                                                if not practica:
                                                    practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        estadosolicitud=3,
                                                        itinerariomalla=itinerariosexto).exists()

                                                if not practica or practicarechazada:
                                                    if not PracticasPreprofesionalesInscripcion.objects.filter(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            actividad__itinerariomalla=itinerariosexto).exists():
                                                        nuevapractica = PracticasPreprofesionalesInscripcion(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                            fechahasta=recordnuevo.materiaregular.nivel.periodo.fin,
                                                            numerohora=isextonuevo.horas_practicas,
                                                            nivelmalla=isextonuevo.nivel,
                                                            tiposolicitud=1,
                                                            estadosolicitud=2,
                                                            tipo=1,
                                                            itinerariomalla=isextonuevo,
                                                            supervisor=profesor,
                                                            tutorunemi=profesor,
                                                            fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                            tipoinstitucion=1,
                                                            sectoreconomico=6,
                                                            empresaempleadora_id=3,
                                                            culminada=True,
                                                            fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                            lugarpractica_id=2,
                                                            observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                        )
                                                        nuevapractica.save()

                                                ItinerariosActividadDetalleDistributivoCarrera.objects.filter(
                                                    status=True,
                                                    itinerario=itinerariosexto).update(
                                                    itinerario=isextonuevo)

                                            if equivalencia.asignaturamallasalto_id == 10953:
                                                itinerarioseptimo = ItinerariosMalla.objects.get(status=True,
                                                                                                 malla_id=mallaantigua_id,
                                                                                                 nivel_id=7)
                                                iseptimonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                             malla_id=mallanueva_id,
                                                                                             nivel_id=7)

                                                practica = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud__in=[
                                                        1, 2, 4,
                                                        5, 6],
                                                    itinerariomalla=itinerarioseptimo).exists()
                                                practicarechazada = False

                                                if not practica:
                                                    practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        estadosolicitud=3,
                                                        itinerariomalla=itinerarioseptimo).exists()

                                                if not practica or practicarechazada:
                                                    if not PracticasPreprofesionalesInscripcion.objects.filter(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            actividad__itinerariomalla=itinerarioseptimo).exists():
                                                        nuevapractica = PracticasPreprofesionalesInscripcion(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                            fechahasta=recordnuevo.materiaregular.nivel.periodo.fin,
                                                            numerohora=iseptimonuevo.horas_practicas,
                                                            nivelmalla=iseptimonuevo.nivel,
                                                            tiposolicitud=1,
                                                            estadosolicitud=2,
                                                            tipo=1,
                                                            itinerariomalla=iseptimonuevo,
                                                            supervisor=profesor,
                                                            tutorunemi=profesor,
                                                            fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                            tipoinstitucion=1,
                                                            sectoreconomico=6,
                                                            empresaempleadora_id=3,
                                                            culminada=True,
                                                            fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                            lugarpractica_id=2,
                                                            observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                        )
                                                        nuevapractica.save()

                                if equivalencia.asignaturamallasalto_id in [10959, 10964]:
                                    if not horasvinculacion:
                                        if recordnuevo.aprobada:
                                            totalhoras = inscripcion.numero_horas_proyectos_vinculacion()
                                            if totalhoras == 0:
                                                a = 0
                                                print('sin horas')
                                            if equivalencia.asignaturamallasalto_id == 10959 and inscripcion.numero_horas_proyectos_vinculacion() < 80:
                                                horasfalta = 80 - inscripcion.numero_horas_proyectos_vinculacion()
                                                vinculacion = ParticipantesMatrices(status=True,
                                                                                    matrizevidencia_id=2,
                                                                                    proyecto_id=601,
                                                                                    inscripcion=inscripcion,
                                                                                    horas=horasfalta,
                                                                                    registrohorasdesde=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                                                    registrohorashasta=recordnuevo.materiaregular.nivel.periodo.fin,
                                                                                    estado=1
                                                                                    )
                                                vinculacion.save()

                                            if equivalencia.asignaturamallasalto_id == 10964 and inscripcion.numero_horas_proyectos_vinculacion() < 160:
                                                horasfalta = 160 - inscripcion.numero_horas_proyectos_vinculacion()
                                                vinculacion = ParticipantesMatrices(status=True,
                                                                                    matrizevidencia_id=2,
                                                                                    proyecto_id=601,
                                                                                    inscripcion=inscripcion,
                                                                                    horas=horasfalta,
                                                                                    registrohorasdesde=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                                                    registrohorashasta=recordnuevo.materiaregular.nivel.periodo.fin,
                                                                                    estado=1
                                                                                    )
                                                vinculacion.save()

                                if not respaldo.exists():
                                    respaldorecord = RespaldoRecordAcademico(
                                        recordacademicooriginal=recordantiguo,
                                        recordacademiconuevo=recordnuevo
                                    )
                                    respaldorecord.save()
                                else:
                                    respaldorecord = respaldo[0]
                                    respaldorecord.recordacademiconuevo = recordnuevo
                                    respaldorecord.save()
                                print(u"Record actualizado %s" % recordnuevo)


                        else:
                            hojadestino.write(fila, 3, "NO ENCONTRO RECORD ANTIGUO %s" % equivalencia.asignaturamalla)
                            fila += 1

                    fila += 1

                    time.sleep(3)

            lin += 1

        libdestino.save(output_folder + libre_origen)
        print(output_folder + libre_origen)
        print("Proceso finalizado. . .")


    except Exception as ex:

        transaction.set_rollback(True)
        print('error: %s' % ex)
        hojadestino.write(fila, 3, str(ex))
        fila += 1

def homologacion_turismo_7():
    try:
        libre_origen = '/homologacion_turismo_2_7.xls'
        fuentecabecera = easyxf(
            'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
        fuentenormal = easyxf(
            'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')

        output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
        # liborigen = xlrd.open_workbook(output_folder + libre_origen)
        libdestino = xlwt.Workbook()
        hojadestino = libdestino.add_sheet('Sheet1')
        fil = 0
        columnas = [(u"CEDULA", 6000),
                    (u"APELLIDOS Y NOMBRES", 6000),
                    (u"OBSERVACIÓN", 6000)
                    ]
        for col_num in range(len(columnas)):
            hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
            hojadestino.col(col_num).width = columnas[col_num][1]
        fila = 1
        titulacion = 0

        lin = 0
        # miarchivo = openpyxl.load_workbook("CEDULA_COMUNICACION_final.xlsx")
        miarchivo = openpyxl.load_workbook("turismolinea.xlsx")
        # miarchivo = openpyxl.load_workbook(path_anexo)
        # miarchivo = openpyxl.load_workbook("prueba_2.xlsx")

        ws = miarchivo.get_sheet_by_name("septimo")
        worksheet = ws
        c = 0
        cont = 0
        periodo_id = 224
        carrera_id = 134
        mallaantigua_id = 199
        mallanueva_id = 487
        for row in worksheet.iter_rows(min_row=0):
            if lin >= 0:
                currentValues, cadena = [], ''
                for cell in row:
                    cadena += str(cell.value) + ' '
                    currentValues.append(str(cell.value))
                identificacion = currentValues[0]

                if not identificacion:
                    break

                with transaction.atomic():
                    matricula = Matricula.objects.filter(status=True, inscripcion__carrera__id=carrera_id,
                                                         inscripcion__persona__cedula=identificacion).first()
                    cont += 1
                    matricula.pasoayuda = True
                    matricula.save()
                    print(u"%s - %s" % (matricula, cont))
                    inscripcion = matricula.inscripcion
                    hojadestino.write(fila, 0, matricula.inscripcion.persona.identificacion(), fuentenormal)
                    hojadestino.write(fila, 1, matricula.inscripcion.persona.nombre_completo(), fuentenormal)
                    hojadestino.write(fila, 2, matricula.nivel.periodo.nombre, fuentenormal)

                    itinerarios = ItinerariosMalla.objects.filter(status=True, malla_id=mallaantigua_id)

                    for itinerario in itinerarios:
                        itinerarionuevo = ItinerariosMalla.objects.filter(status=True, malla_id=mallanueva_id,
                                                                          nivel_id=itinerario.nivel_id)
                        # estadodo culmiado, en curso y pendiente
                        practicasencurso = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                               inscripcion=inscripcion,
                                                                                               estadosolicitud=2,
                                                                                               culminada=False,
                                                                                               itinerariomalla=itinerario)
                        for pcurso in practicasencurso:
                            pcurso.culminada = True
                            pcurso.save()

                        practicaspendientes = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                  inscripcion=inscripcion,
                                                                                                  estadosolicitud=4,
                                                                                                  culminada=False,
                                                                                                  itinerariomalla=itinerario)
                        for pendiente in practicaspendientes:
                            pendiente.estadosolicitud = 2
                            pendiente.culminada = True
                            pendiente.save()

                    practicaspp = haber_cumplido_horas_creditos_practicas_preprofesionales(inscripcion.id)
                    horasvinculacion = haber_cumplido_horas_creditos_vinculacion(inscripcion.id)

                    if InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                       malla_id=mallaantigua_id).exists():
                        imantigua = \
                            InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                            malla_id=mallaantigua_id)[
                                0]
                        imantigua.status = False
                        imantigua.save()
                        print(u"Desactiva antigua inscripcion -----------------------------")

                    if not InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                           malla_id=mallanueva_id).exists():
                        imnueva = InscripcionMalla(inscripcion=inscripcion, malla_id=mallanueva_id)
                        imnueva.save()
                        print(u"Crea nueva inscripcion -----------------------------")

                    equivalencias = TablaEquivalenciaAsignaturas.objects.filter(status=True,
                                                                                asignaturamalla__malla_id=mallaantigua_id).order_by(
                        'asignaturamallasalto__nivelmalla__orden')
                    for equivalencia in equivalencias:
                        print(u"nueva - %s" % equivalencia.asignaturamallasalto)
                        recordantiguo = inscripcion.recordacademico_set.filter(status=True,
                                                                               asignaturamalla_id=equivalencia.asignaturamalla_id).first()

                        if recordantiguo:
                            print(u"anterior - %s" % equivalencia.asignaturamalla)
                            print(u"Record antiguo: %s" % recordantiguo)
                            recordnuevo = None
                            recordantiguo.status = False
                            recordantiguo.save(update_asignaturamalla=False)

                            if equivalencia.asignaturamallasalto_id in [10929, 10947, 10959, 10953, 10964]:
                                observaciones = recordantiguo.observaciones + " Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15"
                                homologada = True



                            else:
                                observaciones = recordantiguo.observaciones + " Migración con base a RESOLUCIÓN CGA-SO-9-2023-NO15 Y RESOLUCIÓN OCS-SO-18-2023-NO14"
                                homologada = False
                            if not RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                  asignaturamalla=equivalencia.asignaturamallasalto).exists():

                                recordnuevo = RecordAcademico(inscripcion=inscripcion,
                                                              matriculas=recordantiguo.matriculas,
                                                              asignaturamalla=equivalencia.asignaturamallasalto,
                                                              asignatura=equivalencia.asignaturamallasalto.asignatura,
                                                              asignaturaold_id=recordantiguo.asignatura.id,
                                                              nota=recordantiguo.nota,
                                                              asistencia=recordantiguo.asistencia,
                                                              sinasistencia=recordantiguo.sinasistencia,
                                                              fecha=recordantiguo.fecha,
                                                              noaplica=recordantiguo.noaplica,
                                                              aprobada=recordantiguo.aprobada,
                                                              convalidacion=recordantiguo.convalidacion,
                                                              pendiente=recordantiguo.pendiente,
                                                              creditos=equivalencia.asignaturamallasalto.creditos,
                                                              horas=equivalencia.asignaturamallasalto.horas,
                                                              valida=recordantiguo.valida,
                                                              validapromedio=recordantiguo.validapromedio,
                                                              observaciones=observaciones,
                                                              homologada=homologada,
                                                              materiaregular=recordantiguo.materiaregular,
                                                              materiacurso=None,
                                                              completonota=recordantiguo.completonota,
                                                              completoasistencia=recordantiguo.completoasistencia,
                                                              fechainicio=recordantiguo.fechainicio,
                                                              fechafin=recordantiguo.fechafin,
                                                              suficiencia=recordantiguo.suficiencia,
                                                              asignaturamallahistorico_id=recordantiguo.asignaturamalla.id,
                                                              reverso=False)
                                recordnuevo.save()
                                print(u"Crea nuevo record %s" % recordnuevo)


                            elif RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                asignaturamalla=equivalencia.asignaturamallasalto):
                                recordnuevo = RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                             asignaturamalla=equivalencia.asignaturamallasalto)[
                                    0]
                                recordnuevo.matriculas = recordantiguo.matriculas
                                recordnuevo.asignaturamalla = equivalencia.asignaturamallasalto
                                recordnuevo.asignatura = equivalencia.asignaturamallasalto.asignatura
                                recordnuevo.asignaturaold = recordantiguo.asignaturamalla.asignatura
                                recordnuevo.nota = recordantiguo.nota
                                recordnuevo.asistencia = recordantiguo.asistencia
                                recordnuevo.sinasistencia = recordantiguo.sinasistencia
                                recordnuevo.fecha = recordantiguo.fecha
                                recordnuevo.noaplica = recordantiguo.noaplica
                                recordnuevo.aprobada = recordantiguo.aprobada
                                recordnuevo.convalidacion = recordantiguo.convalidacion
                                recordnuevo.pendiente = recordantiguo.pendiente
                                recordnuevo.creditos = equivalencia.asignaturamallasalto.creditos
                                recordnuevo.horas = equivalencia.asignaturamallasalto.horas
                                recordnuevo.valida = recordantiguo.valida
                                recordnuevo.validapromedio = recordantiguo.validapromedio
                                recordnuevo.observaciones = observaciones
                                recordnuevo.homologada = homologada
                                recordnuevo.materiaregular = recordantiguo.materiaregular
                                recordnuevo.materiacurso = None
                                recordnuevo.completonota = recordantiguo.completonota
                                recordnuevo.completoasistencia = recordantiguo.completoasistencia
                                recordnuevo.fechainicio = recordantiguo.fechainicio
                                recordnuevo.fechafin = recordantiguo.fechafin
                                recordnuevo.suficiencia = recordantiguo.suficiencia
                                recordnuevo.asignaturamallahistorico = recordantiguo.asignaturamalla
                                recordnuevo.reverso = False
                                recordnuevo.save()

                            if recordnuevo:
                                historicos = HistoricoRecordAcademico.objects.filter(status=True,
                                                                                     recordacademico=recordantiguo).update(
                                    recordacademico=recordnuevo,
                                    creditos=recordnuevo.creditos,
                                    horas=recordnuevo.horas,
                                    homologada=recordnuevo.homologada)
                                respaldo = RespaldoRecordAcademico.objects.filter(status=True,
                                                                                  recordacademicooriginal=recordantiguo)

                                if equivalencia.asignaturamallasalto_id in [10929, 10947, 10953]:
                                    if not practicaspp:
                                        if recordnuevo.aprobada:
                                            profesor = None
                                            if recordnuevo.materiaregular:
                                                profesor = recordnuevo.materiaregular.profesor_principal()
                                            elif recordnuevo.materiacurso:
                                                profesor = recordnuevo.materiaregular.profesor()
                                            if equivalencia.asignaturamallasalto_id == 10929:
                                                itinerarioquinto = ItinerariosMalla.objects.get(status=True,
                                                                                                malla_id=mallaantigua_id,
                                                                                                nivel_id=5)
                                                iquintonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                            malla_id=mallanueva_id,
                                                                                            nivel_id=5)

                                                practica = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud__in=[
                                                        1, 2, 4,
                                                        5, 6],
                                                    itinerariomalla=itinerarioquinto).exists()
                                                practicarechazada = False

                                                if not practica:
                                                    practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        estadosolicitud=3,
                                                        itinerariomalla=itinerarioquinto).exists()

                                                if not practica or practicarechazada:
                                                    if not PracticasPreprofesionalesInscripcion.objects.filter(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            actividad__itinerariomalla=itinerarioquinto).exists():
                                                        nuevapractica = PracticasPreprofesionalesInscripcion(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            numerohora=iquintonuevo.horas_practicas,
                                                            nivelmalla=iquintonuevo.nivel,
                                                            tiposolicitud=1,
                                                            estadosolicitud=2,
                                                            tipo=1,
                                                            itinerariomalla=iquintonuevo,
                                                            supervisor=profesor,
                                                            tutorunemi=profesor,
                                                            fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            tipoinstitucion=1,
                                                            sectoreconomico=6,
                                                            empresaempleadora_id=3,
                                                            culminada=True,
                                                            fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            lugarpractica_id=2,
                                                            observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                        )
                                                        nuevapractica.save()

                                                ItinerariosActividadDetalleDistributivoCarrera.objects.filter(
                                                    status=True,
                                                    itinerario=itinerarioquinto).update(
                                                    itinerario=iquintonuevo)

                                            if equivalencia.asignaturamallasalto_id == 10947:
                                                itinerariosexto = ItinerariosMalla.objects.get(status=True,
                                                                                               malla_id=mallaantigua_id,
                                                                                               nivel_id=6)
                                                isextonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallanueva_id,
                                                                                           nivel_id=6)

                                                practica = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud__in=[
                                                        1, 2, 4,
                                                        5, 6],
                                                    itinerariomalla=itinerariosexto).exists()
                                                practicarechazada = False

                                                if not practica:
                                                    practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        estadosolicitud=3,
                                                        itinerariomalla=itinerariosexto).exists()

                                                if not practica or practicarechazada:
                                                    if not PracticasPreprofesionalesInscripcion.objects.filter(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            actividad__itinerariomalla=itinerariosexto).exists():
                                                        nuevapractica = PracticasPreprofesionalesInscripcion(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                            fechahasta=recordnuevo.materiaregular.nivel.periodo.fin,
                                                            numerohora=isextonuevo.horas_practicas,
                                                            nivelmalla=isextonuevo.nivel,
                                                            tiposolicitud=1,
                                                            estadosolicitud=2,
                                                            tipo=1,
                                                            itinerariomalla=isextonuevo,
                                                            supervisor=profesor,
                                                            tutorunemi=profesor,
                                                            fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                            tipoinstitucion=1,
                                                            sectoreconomico=6,
                                                            empresaempleadora_id=3,
                                                            culminada=True,
                                                            fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                            lugarpractica_id=2,
                                                            observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                        )
                                                        nuevapractica.save()

                                                ItinerariosActividadDetalleDistributivoCarrera.objects.filter(
                                                    status=True,
                                                    itinerario=itinerariosexto).update(
                                                    itinerario=isextonuevo)

                                            if equivalencia.asignaturamallasalto_id == 10953:
                                                itinerarioseptimo = ItinerariosMalla.objects.get(status=True,
                                                                                                 malla_id=mallaantigua_id,
                                                                                                 nivel_id=7)
                                                iseptimonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                             malla_id=mallanueva_id,
                                                                                             nivel_id=7)

                                                practica = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud__in=[
                                                        1, 2, 4,
                                                        5, 6],
                                                    itinerariomalla=itinerarioseptimo).exists()
                                                practicarechazada = False

                                                if not practica:
                                                    practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        estadosolicitud=3,
                                                        itinerariomalla=itinerarioseptimo).exists()

                                                if not practica or practicarechazada:
                                                    if not PracticasPreprofesionalesInscripcion.objects.filter(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            actividad__itinerariomalla=itinerarioseptimo).exists():
                                                        nuevapractica = PracticasPreprofesionalesInscripcion(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                            fechahasta=recordnuevo.materiaregular.nivel.periodo.fin,
                                                            numerohora=iseptimonuevo.horas_practicas,
                                                            nivelmalla=iseptimonuevo.nivel,
                                                            tiposolicitud=1,
                                                            estadosolicitud=2,
                                                            tipo=1,
                                                            itinerariomalla=iseptimonuevo,
                                                            supervisor=profesor,
                                                            tutorunemi=profesor,
                                                            fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                            tipoinstitucion=1,
                                                            sectoreconomico=6,
                                                            empresaempleadora_id=3,
                                                            culminada=True,
                                                            fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                            lugarpractica_id=2,
                                                            observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                        )
                                                        nuevapractica.save()

                                if equivalencia.asignaturamallasalto_id in [10959, 10964]:
                                    if not horasvinculacion:
                                        if recordnuevo.aprobada:
                                            totalhoras = inscripcion.numero_horas_proyectos_vinculacion()
                                            if totalhoras == 0:
                                                a = 0
                                                print('sin horas')
                                            if equivalencia.asignaturamallasalto_id == 10959 and inscripcion.numero_horas_proyectos_vinculacion() < 80:
                                                horasfalta = 80 - inscripcion.numero_horas_proyectos_vinculacion()
                                                vinculacion = ParticipantesMatrices(status=True,
                                                                                    matrizevidencia_id=2,
                                                                                    proyecto_id=601,
                                                                                    inscripcion=inscripcion,
                                                                                    horas=horasfalta,
                                                                                    registrohorasdesde=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                                                    registrohorashasta=recordnuevo.materiaregular.nivel.periodo.fin,
                                                                                    estado=1
                                                                                    )
                                                vinculacion.save()

                                            if equivalencia.asignaturamallasalto_id == 10964 and inscripcion.numero_horas_proyectos_vinculacion() < 160:
                                                horasfalta = 160 - inscripcion.numero_horas_proyectos_vinculacion()
                                                vinculacion = ParticipantesMatrices(status=True,
                                                                                    matrizevidencia_id=2,
                                                                                    proyecto_id=601,
                                                                                    inscripcion=inscripcion,
                                                                                    horas=horasfalta,
                                                                                    registrohorasdesde=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                                                    registrohorashasta=recordnuevo.materiaregular.nivel.periodo.fin,
                                                                                    estado=1
                                                                                    )
                                                vinculacion.save()

                                if not respaldo.exists():
                                    respaldorecord = RespaldoRecordAcademico(
                                        recordacademicooriginal=recordantiguo,
                                        recordacademiconuevo=recordnuevo
                                    )
                                    respaldorecord.save()
                                else:
                                    respaldorecord = respaldo[0]
                                    respaldorecord.recordacademiconuevo = recordnuevo
                                    respaldorecord.save()
                                print(u"Record actualizado %s" % recordnuevo)


                        else:
                            hojadestino.write(fila, 3, "NO ENCONTRO RECORD ANTIGUO %s" % equivalencia.asignaturamalla)
                            fila += 1

                    fila += 1

                    time.sleep(3)

            lin += 1

        libdestino.save(output_folder + libre_origen)
        print(output_folder + libre_origen)
        print("Proceso finalizado. . .")


    except Exception as ex:

        transaction.set_rollback(True)
        print('error: %s' % ex)
        hojadestino.write(fila, 3, str(ex))
        fila += 1

def homologacion_turismo_8():
    try:
        libre_origen = '/homologacion_turismo_2_8.xls'
        fuentecabecera = easyxf(
            'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
        fuentenormal = easyxf(
            'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')

        output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
        # liborigen = xlrd.open_workbook(output_folder + libre_origen)
        libdestino = xlwt.Workbook()
        hojadestino = libdestino.add_sheet('Sheet1')
        fil = 0
        columnas = [(u"CEDULA", 6000),
                    (u"APELLIDOS Y NOMBRES", 6000),
                    (u"OBSERVACIÓN", 6000)
                    ]
        for col_num in range(len(columnas)):
            hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
            hojadestino.col(col_num).width = columnas[col_num][1]
        fila = 1
        titulacion = 0

        lin = 0
        # miarchivo = openpyxl.load_workbook("CEDULA_COMUNICACION_final.xlsx")
        miarchivo = openpyxl.load_workbook("turismolinea.xlsx")
        # miarchivo = openpyxl.load_workbook(path_anexo)
        # miarchivo = openpyxl.load_workbook("prueba_2.xlsx")

        ws = miarchivo.get_sheet_by_name("octavo")
        worksheet = ws
        c = 0
        cont = 0
        periodo_id = 224
        carrera_id = 134
        mallaantigua_id = 199
        mallanueva_id = 487
        for row in worksheet.iter_rows(min_row=0):
            if lin >= 0:
                currentValues, cadena = [], ''
                for cell in row:
                    cadena += str(cell.value) + ' '
                    currentValues.append(str(cell.value))
                identificacion = currentValues[0]

                if not identificacion:
                    break

                with transaction.atomic():
                    matricula = Matricula.objects.filter(status=True, inscripcion__carrera__id=carrera_id,
                                                         inscripcion__persona__cedula=identificacion).first()
                    cont += 1
                    matricula.pasoayuda = True
                    matricula.save()
                    print(u"%s - %s" % (matricula, cont))
                    inscripcion = matricula.inscripcion
                    hojadestino.write(fila, 0, matricula.inscripcion.persona.identificacion(), fuentenormal)
                    hojadestino.write(fila, 1, matricula.inscripcion.persona.nombre_completo(), fuentenormal)
                    hojadestino.write(fila, 2, matricula.nivel.periodo.nombre, fuentenormal)

                    itinerarios = ItinerariosMalla.objects.filter(status=True, malla_id=mallaantigua_id)

                    for itinerario in itinerarios:
                        itinerarionuevo = ItinerariosMalla.objects.filter(status=True, malla_id=mallanueva_id,
                                                                          nivel_id=itinerario.nivel_id)
                        # estadodo culmiado, en curso y pendiente
                        practicasencurso = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                               inscripcion=inscripcion,
                                                                                               estadosolicitud=2,
                                                                                               culminada=False,
                                                                                               itinerariomalla=itinerario)
                        for pcurso in practicasencurso:
                            pcurso.culminada = True
                            pcurso.save()

                        practicaspendientes = PracticasPreprofesionalesInscripcion.objects.filter(status=True,
                                                                                                  inscripcion=inscripcion,
                                                                                                  estadosolicitud=4,
                                                                                                  culminada=False,
                                                                                                  itinerariomalla=itinerario)
                        for pendiente in practicaspendientes:
                            pendiente.estadosolicitud = 2
                            pendiente.culminada = True
                            pendiente.save()

                    practicaspp = haber_cumplido_horas_creditos_practicas_preprofesionales(inscripcion.id)
                    horasvinculacion = haber_cumplido_horas_creditos_vinculacion(inscripcion.id)

                    if InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                       malla_id=mallaantigua_id).exists():
                        imantigua = \
                            InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                            malla_id=mallaantigua_id)[
                                0]
                        imantigua.status = False
                        imantigua.save()
                        print(u"Desactiva antigua inscripcion -----------------------------")

                    if not InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                           malla_id=mallanueva_id).exists():
                        imnueva = InscripcionMalla(inscripcion=inscripcion, malla_id=mallanueva_id)
                        imnueva.save()
                        print(u"Crea nueva inscripcion -----------------------------")

                    equivalencias = TablaEquivalenciaAsignaturas.objects.filter(status=True,
                                                                                asignaturamalla__malla_id=mallaantigua_id).order_by(
                        'asignaturamallasalto__nivelmalla__orden')
                    for equivalencia in equivalencias:
                        print(u"nueva - %s" % equivalencia.asignaturamallasalto)
                        recordantiguo = inscripcion.recordacademico_set.filter(status=True,
                                                                               asignaturamalla_id=equivalencia.asignaturamalla_id).first()

                        if recordantiguo:
                            print(u"anterior - %s" % equivalencia.asignaturamalla)
                            print(u"Record antiguo: %s" % recordantiguo)
                            recordnuevo = None
                            recordantiguo.status = False
                            recordantiguo.save(update_asignaturamalla=False)

                            if equivalencia.asignaturamallasalto_id in [10929, 10947, 10959, 10953, 10964]:
                                observaciones = recordantiguo.observaciones + " Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15"
                                homologada = True



                            else:
                                observaciones = recordantiguo.observaciones + " Migración con base a RESOLUCIÓN CGA-SO-9-2023-NO15 Y RESOLUCIÓN OCS-SO-18-2023-NO14"
                                homologada = False
                            if not RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                  asignaturamalla=equivalencia.asignaturamallasalto).exists():

                                recordnuevo = RecordAcademico(inscripcion=inscripcion,
                                                              matriculas=recordantiguo.matriculas,
                                                              asignaturamalla=equivalencia.asignaturamallasalto,
                                                              asignatura=equivalencia.asignaturamallasalto.asignatura,
                                                              asignaturaold_id=recordantiguo.asignatura.id,
                                                              nota=recordantiguo.nota,
                                                              asistencia=recordantiguo.asistencia,
                                                              sinasistencia=recordantiguo.sinasistencia,
                                                              fecha=recordantiguo.fecha,
                                                              noaplica=recordantiguo.noaplica,
                                                              aprobada=recordantiguo.aprobada,
                                                              convalidacion=recordantiguo.convalidacion,
                                                              pendiente=recordantiguo.pendiente,
                                                              creditos=equivalencia.asignaturamallasalto.creditos,
                                                              horas=equivalencia.asignaturamallasalto.horas,
                                                              valida=recordantiguo.valida,
                                                              validapromedio=recordantiguo.validapromedio,
                                                              observaciones=observaciones,
                                                              homologada=homologada,
                                                              materiaregular=recordantiguo.materiaregular,
                                                              materiacurso=None,
                                                              completonota=recordantiguo.completonota,
                                                              completoasistencia=recordantiguo.completoasistencia,
                                                              fechainicio=recordantiguo.fechainicio,
                                                              fechafin=recordantiguo.fechafin,
                                                              suficiencia=recordantiguo.suficiencia,
                                                              asignaturamallahistorico_id=recordantiguo.asignaturamalla.id,
                                                              reverso=False)
                                recordnuevo.save()
                                print(u"Crea nuevo record %s" % recordnuevo)


                            elif RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                asignaturamalla=equivalencia.asignaturamallasalto):
                                recordnuevo = RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                             asignaturamalla=equivalencia.asignaturamallasalto)[
                                    0]
                                recordnuevo.matriculas = recordantiguo.matriculas
                                recordnuevo.asignaturamalla = equivalencia.asignaturamallasalto
                                recordnuevo.asignatura = equivalencia.asignaturamallasalto.asignatura
                                recordnuevo.asignaturaold = recordantiguo.asignaturamalla.asignatura
                                recordnuevo.nota = recordantiguo.nota
                                recordnuevo.asistencia = recordantiguo.asistencia
                                recordnuevo.sinasistencia = recordantiguo.sinasistencia
                                recordnuevo.fecha = recordantiguo.fecha
                                recordnuevo.noaplica = recordantiguo.noaplica
                                recordnuevo.aprobada = recordantiguo.aprobada
                                recordnuevo.convalidacion = recordantiguo.convalidacion
                                recordnuevo.pendiente = recordantiguo.pendiente
                                recordnuevo.creditos = equivalencia.asignaturamallasalto.creditos
                                recordnuevo.horas = equivalencia.asignaturamallasalto.horas
                                recordnuevo.valida = recordantiguo.valida
                                recordnuevo.validapromedio = recordantiguo.validapromedio
                                recordnuevo.observaciones = observaciones
                                recordnuevo.homologada = homologada
                                recordnuevo.materiaregular = recordantiguo.materiaregular
                                recordnuevo.materiacurso = None
                                recordnuevo.completonota = recordantiguo.completonota
                                recordnuevo.completoasistencia = recordantiguo.completoasistencia
                                recordnuevo.fechainicio = recordantiguo.fechainicio
                                recordnuevo.fechafin = recordantiguo.fechafin
                                recordnuevo.suficiencia = recordantiguo.suficiencia
                                recordnuevo.asignaturamallahistorico = recordantiguo.asignaturamalla
                                recordnuevo.reverso = False
                                recordnuevo.save()

                            if recordnuevo:
                                historicos = HistoricoRecordAcademico.objects.filter(status=True,
                                                                                     recordacademico=recordantiguo).update(
                                    recordacademico=recordnuevo,
                                    creditos=recordnuevo.creditos,
                                    horas=recordnuevo.horas,
                                    homologada=recordnuevo.homologada)
                                respaldo = RespaldoRecordAcademico.objects.filter(status=True,
                                                                                  recordacademicooriginal=recordantiguo)

                                if equivalencia.asignaturamallasalto_id in [10929, 10947, 10953]:
                                    if not practicaspp:
                                        if recordnuevo.aprobada:
                                            profesor = None
                                            if recordnuevo.materiaregular:
                                                profesor = recordnuevo.materiaregular.profesor_principal()
                                            elif recordnuevo.materiacurso:
                                                profesor = recordnuevo.materiaregular.profesor()
                                            if equivalencia.asignaturamallasalto_id == 10929:
                                                itinerarioquinto = ItinerariosMalla.objects.get(status=True,
                                                                                                malla_id=mallaantigua_id,
                                                                                                nivel_id=5)
                                                iquintonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                            malla_id=mallanueva_id,
                                                                                            nivel_id=5)

                                                practica = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud__in=[
                                                        1, 2, 4,
                                                        5, 6],
                                                    itinerariomalla=itinerarioquinto).exists()
                                                practicarechazada = False

                                                if not practica:
                                                    practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        estadosolicitud=3,
                                                        itinerariomalla=itinerarioquinto).exists()

                                                if not practica or practicarechazada:
                                                    if not PracticasPreprofesionalesInscripcion.objects.filter(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            actividad__itinerariomalla=itinerarioquinto).exists():
                                                        nuevapractica = PracticasPreprofesionalesInscripcion(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            fechahasta=recordnuevo.materiaregular.nivel.periodo.fin if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            numerohora=iquintonuevo.horas_practicas,
                                                            nivelmalla=iquintonuevo.nivel,
                                                            tiposolicitud=1,
                                                            estadosolicitud=2,
                                                            tipo=1,
                                                            itinerariomalla=iquintonuevo,
                                                            supervisor=profesor,
                                                            tutorunemi=profesor,
                                                            fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            tipoinstitucion=1,
                                                            sectoreconomico=6,
                                                            empresaempleadora_id=3,
                                                            culminada=True,
                                                            fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio if not recordnuevo.materiaregular == None else datetime.now().date(),
                                                            lugarpractica_id=2,
                                                            observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                        )
                                                        nuevapractica.save()

                                                ItinerariosActividadDetalleDistributivoCarrera.objects.filter(
                                                    status=True,
                                                    itinerario=itinerarioquinto).update(
                                                    itinerario=iquintonuevo)

                                            if equivalencia.asignaturamallasalto_id == 10947:
                                                itinerariosexto = ItinerariosMalla.objects.get(status=True,
                                                                                               malla_id=mallaantigua_id,
                                                                                               nivel_id=6)
                                                isextonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                           malla_id=mallanueva_id,
                                                                                           nivel_id=6)

                                                practica = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud__in=[
                                                        1, 2, 4,
                                                        5, 6],
                                                    itinerariomalla=itinerariosexto).exists()
                                                practicarechazada = False

                                                if not practica:
                                                    practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        estadosolicitud=3,
                                                        itinerariomalla=itinerariosexto).exists()

                                                if not practica or practicarechazada:
                                                    if not PracticasPreprofesionalesInscripcion.objects.filter(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            actividad__itinerariomalla=itinerariosexto).exists():
                                                        nuevapractica = PracticasPreprofesionalesInscripcion(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                            fechahasta=recordnuevo.materiaregular.nivel.periodo.fin,
                                                            numerohora=isextonuevo.horas_practicas,
                                                            nivelmalla=isextonuevo.nivel,
                                                            tiposolicitud=1,
                                                            estadosolicitud=2,
                                                            tipo=1,
                                                            itinerariomalla=isextonuevo,
                                                            supervisor=profesor,
                                                            tutorunemi=profesor,
                                                            fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                            tipoinstitucion=1,
                                                            sectoreconomico=6,
                                                            empresaempleadora_id=3,
                                                            culminada=True,
                                                            fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                            lugarpractica_id=2,
                                                            observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                        )
                                                        nuevapractica.save()

                                                ItinerariosActividadDetalleDistributivoCarrera.objects.filter(
                                                    status=True,
                                                    itinerario=itinerariosexto).update(
                                                    itinerario=isextonuevo)

                                            if equivalencia.asignaturamallasalto_id == 10953:
                                                itinerarioseptimo = ItinerariosMalla.objects.get(status=True,
                                                                                                 malla_id=mallaantigua_id,
                                                                                                 nivel_id=7)
                                                iseptimonuevo = ItinerariosMalla.objects.get(status=True,
                                                                                             malla_id=mallanueva_id,
                                                                                             nivel_id=7)

                                                practica = PracticasPreprofesionalesInscripcion.objects.filter(
                                                    status=True,
                                                    inscripcion=inscripcion,
                                                    estadosolicitud__in=[
                                                        1, 2, 4,
                                                        5, 6],
                                                    itinerariomalla=itinerarioseptimo).exists()
                                                practicarechazada = False

                                                if not practica:
                                                    practicarechazada = PracticasPreprofesionalesInscripcion.objects.filter(
                                                        status=True,
                                                        inscripcion=inscripcion,
                                                        estadosolicitud=3,
                                                        itinerariomalla=itinerarioseptimo).exists()

                                                if not practica or practicarechazada:
                                                    if not PracticasPreprofesionalesInscripcion.objects.filter(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            actividad__itinerariomalla=itinerarioseptimo).exists():
                                                        nuevapractica = PracticasPreprofesionalesInscripcion(
                                                            status=True,
                                                            inscripcion=inscripcion,
                                                            fechadesde=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                            fechahasta=recordnuevo.materiaregular.nivel.periodo.fin,
                                                            numerohora=iseptimonuevo.horas_practicas,
                                                            nivelmalla=iseptimonuevo.nivel,
                                                            tiposolicitud=1,
                                                            estadosolicitud=2,
                                                            tipo=1,
                                                            itinerariomalla=iseptimonuevo,
                                                            supervisor=profesor,
                                                            tutorunemi=profesor,
                                                            fechaasigtutor=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                            tipoinstitucion=1,
                                                            sectoreconomico=6,
                                                            empresaempleadora_id=3,
                                                            culminada=True,
                                                            fechaasigsupervisor=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                            lugarpractica_id=2,
                                                            observacion='Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15'
                                                        )
                                                        nuevapractica.save()

                                if equivalencia.asignaturamallasalto_id in [10959, 10964]:
                                    if not horasvinculacion:
                                        if recordnuevo.aprobada:
                                            totalhoras = inscripcion.numero_horas_proyectos_vinculacion()
                                            if totalhoras == 0:
                                                a = 0
                                                print('sin horas')
                                            if equivalencia.asignaturamallasalto_id == 10959 and inscripcion.numero_horas_proyectos_vinculacion() < 80:
                                                horasfalta = 80 - inscripcion.numero_horas_proyectos_vinculacion()
                                                vinculacion = ParticipantesMatrices(status=True,
                                                                                    matrizevidencia_id=2,
                                                                                    proyecto_id=601,
                                                                                    inscripcion=inscripcion,
                                                                                    horas=horasfalta,
                                                                                    registrohorasdesde=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                                                    registrohorashasta=recordnuevo.materiaregular.nivel.periodo.fin,
                                                                                    estado=1
                                                                                    )
                                                vinculacion.save()

                                            if equivalencia.asignaturamallasalto_id == 10964 and inscripcion.numero_horas_proyectos_vinculacion() < 160:
                                                horasfalta = 160 - inscripcion.numero_horas_proyectos_vinculacion()
                                                vinculacion = ParticipantesMatrices(status=True,
                                                                                    matrizevidencia_id=2,
                                                                                    proyecto_id=601,
                                                                                    inscripcion=inscripcion,
                                                                                    horas=horasfalta,
                                                                                    registrohorasdesde=recordnuevo.materiaregular.nivel.periodo.inicio,
                                                                                    registrohorashasta=recordnuevo.materiaregular.nivel.periodo.fin,
                                                                                    estado=1
                                                                                    )
                                                vinculacion.save()

                                if not respaldo.exists():
                                    respaldorecord = RespaldoRecordAcademico(
                                        recordacademicooriginal=recordantiguo,
                                        recordacademiconuevo=recordnuevo
                                    )
                                    respaldorecord.save()
                                else:
                                    respaldorecord = respaldo[0]
                                    respaldorecord.recordacademiconuevo = recordnuevo
                                    respaldorecord.save()
                                print(u"Record actualizado %s" % recordnuevo)


                        else:
                            hojadestino.write(fila, 3, "NO ENCONTRO RECORD ANTIGUO %s" % equivalencia.asignaturamalla)
                            fila += 1

                    fila += 1

                    time.sleep(3)

            lin += 1

        libdestino.save(output_folder + libre_origen)
        print(output_folder + libre_origen)
        print("Proceso finalizado. . .")


    except Exception as ex:

        transaction.set_rollback(True)
        print('error: %s' % ex)
        hojadestino.write(fila, 3, str(ex))
        fila += 1


def homologacion_derecho_1():

    #verificar que todos los estudiantes tengan la misma malla
    cadena = ''
    linea, excluidos, conexito = 0, 0, 0
    try:
        libre_origen = '/homologacion_derecho_2_1.xls'
        fuentecabecera = easyxf(
            'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
        fuentenormal = easyxf(
            'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
        output_folder = MEDIA_ROOT
        output_folder = os.path.join(os.path.join(BASE_DIR))
        # liborigen = xlrd.open_workbook(output_folder + libre_origen)
        libdestino = xlwt.Workbook()
        hojadestino = libdestino.add_sheet('HOJA1')
        fil = 0
        columnas = [(u"CEDULA", 6000),
                   (u"APELLIDOS Y NOMBRES", 6000),
                   (u"OBSERVACIÓN", 6000)
        ]
        for col_num in range(len(columnas)):
            hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
            hojadestino.col(col_num).width = columnas[col_num][1]
        fila = 1
        titulacion = 0

        lin = 0
        miarchivo = openpyxl.load_workbook("derecholineaf.xlsx")
        #miarchivo = openpyxl.load_workbook("prueba_2.xlsx")

        ws = miarchivo.get_sheet_by_name("primero")
        worksheet = ws
        c = 0
        cont = 0
        periodo_id=177
        carrera_id=126
        mallaantigua_id=207
        mallanueva_id=480

        for row in worksheet.iter_rows(min_row=0):
            if lin >= 0:
                currentValues, cadena = [], ''
                for cell in row:
                    cadena += str(cell.value) + ' '
                    currentValues.append(str(cell.value))
                identificacion = currentValues[0]

                with transaction.atomic():
                    matricula = Matricula.objects.filter(status=True, inscripcion__carrera__id=carrera_id, inscripcion__persona__cedula=identificacion).first()
                    cont += 1
                    matricula.pasoayuda = True
                    matricula.save()
                    print(u"%s - %s" % (matricula, cont))
                    inscripcion = matricula.inscripcion
                    hojadestino.write(fila, 0, matricula.inscripcion.persona.identificacion(), fuentenormal)
                    hojadestino.write(fila, 1, matricula.inscripcion.persona.nombre_completo(), fuentenormal)
                    hojadestino.write(fila, 2, matricula.nivel.periodo.nombre, fuentenormal)

                    if InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallaantigua_id).exists():
                        imantigua = InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallaantigua_id)[0]
                        imantigua.status = False
                        imantigua.save()
                        print(u"Desactiva antigua inscripcion -----------------------------")

                    if not InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallanueva_id).exists():
                        imnueva = InscripcionMalla(inscripcion=inscripcion, malla_id=mallanueva_id)
                        imnueva.save()
                        print(u"Crea nueva inscripcion -----------------------------")

                    equivalencias = TablaEquivalenciaAsignaturas.objects.filter(status=True, asignaturamalla__malla_id=mallaantigua_id).order_by('asignaturamallasalto__nivelmalla__orden')
                    for equivalencia in equivalencias:
                        print(u"nueva - %s" % equivalencia.asignaturamallasalto)
                        recordantiguo = inscripcion.recordacademico_set.filter(status=True,asignaturamalla_id=equivalencia.asignaturamalla_id).first()

                        if recordantiguo:
                            print(u"anterior - %s" % equivalencia.asignaturamalla)
                            print(u"Record antiguo: %s" % recordantiguo)
                            recordnuevo = None
                            recordantiguo.status = False
                            recordantiguo.save(update_asignaturamalla=False)

                            if equivalencia.asignaturamallasalto_id == 10623 or equivalencia.asignaturamallasalto_id == 10627 or equivalencia.asignaturamallasalto_id == 10615:
                                observaciones = recordantiguo.observaciones + " Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15"
                                homologada = True
                            else:
                                observaciones = recordantiguo.observaciones + " Migración con base a RESOLUCIÓN CGA-SO-9-2023-NO15 Y RESOLUCIÓN OCS-SO-18-2023-NO14"
                                homologada = False
                            if not RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                  asignaturamalla=equivalencia.asignaturamallasalto).exists():



                                recordnuevo = RecordAcademico(inscripcion=inscripcion,
                                                              matriculas=recordantiguo.matriculas,
                                                              asignaturamalla=equivalencia.asignaturamallasalto,
                                                              asignatura=equivalencia.asignaturamallasalto.asignatura,
                                                              asignaturaold_id=recordantiguo.asignatura.id,
                                                              nota=recordantiguo.nota,
                                                              asistencia=recordantiguo.asistencia,
                                                              sinasistencia=recordantiguo.sinasistencia,
                                                              fecha=recordantiguo.fecha,
                                                              noaplica=recordantiguo.noaplica,
                                                              aprobada=recordantiguo.aprobada,
                                                              convalidacion=recordantiguo.convalidacion,
                                                              pendiente=recordantiguo.pendiente,
                                                              creditos=equivalencia.asignaturamallasalto.creditos,
                                                              horas=equivalencia.asignaturamallasalto.horas,
                                                              valida=recordantiguo.valida,
                                                              validapromedio=recordantiguo.validapromedio,
                                                              observaciones=observaciones,
                                                              homologada=homologada,
                                                              materiaregular=recordantiguo.materiaregular,
                                                              materiacurso=None,
                                                              completonota=recordantiguo.completonota,
                                                              completoasistencia=recordantiguo.completoasistencia,
                                                              fechainicio=recordantiguo.fechainicio,
                                                              fechafin=recordantiguo.fechafin,
                                                              suficiencia=recordantiguo.suficiencia,
                                                              asignaturamallahistorico_id=recordantiguo.asignaturamalla.id,
                                                              reverso=False)
                                recordnuevo.save()
                                print(u"Crea nuevo record %s" % recordnuevo)


                            elif RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                asignaturamalla=equivalencia.asignaturamallasalto):
                                recordnuevo = RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                             asignaturamalla=equivalencia.asignaturamallasalto)[0]
                                recordnuevo.matriculas = recordantiguo.matriculas
                                recordnuevo.asignaturamalla = equivalencia.asignaturamallasalto
                                recordnuevo.asignatura = equivalencia.asignaturamallasalto.asignatura
                                recordnuevo.asignaturaold = recordantiguo.asignaturamalla.asignatura
                                recordnuevo.nota = recordantiguo.nota
                                recordnuevo.asistencia = recordantiguo.asistencia
                                recordnuevo.sinasistencia = recordantiguo.sinasistencia
                                recordnuevo.fecha = recordantiguo.fecha
                                recordnuevo.noaplica = recordantiguo.noaplica
                                recordnuevo.aprobada = recordantiguo.aprobada
                                recordnuevo.convalidacion = recordantiguo.convalidacion
                                recordnuevo.pendiente = recordantiguo.pendiente
                                recordnuevo.creditos = equivalencia.asignaturamallasalto.creditos
                                recordnuevo.horas = equivalencia.asignaturamallasalto.horas
                                recordnuevo.valida = recordantiguo.valida
                                recordnuevo.validapromedio = recordantiguo.validapromedio
                                recordnuevo.observaciones = observaciones
                                recordnuevo.homologada = homologada
                                recordnuevo.materiaregular = recordantiguo.materiaregular
                                recordnuevo.materiacurso = None
                                recordnuevo.completonota = recordantiguo.completonota
                                recordnuevo.completoasistencia = recordantiguo.completoasistencia
                                recordnuevo.fechainicio = recordantiguo.fechainicio
                                recordnuevo.fechafin = recordantiguo.fechafin
                                recordnuevo.suficiencia = recordantiguo.suficiencia
                                recordnuevo.asignaturamallahistorico = recordantiguo.asignaturamalla
                                recordnuevo.reverso = False
                                recordnuevo.save()

                            if recordnuevo:
                                historicos = HistoricoRecordAcademico.objects.filter(status=True,
                                                                                     recordacademico=recordantiguo).update(recordacademico=recordnuevo,
                                                                                                                           creditos=recordnuevo.creditos,
                                                                                                                           horas=recordnuevo.horas,
                                                                                                                           homologada=recordnuevo.homologada)
                                respaldo = RespaldoRecordAcademico.objects.filter(status=True,recordacademicooriginal=recordantiguo)

                                if not respaldo.exists():
                                    respaldorecord = RespaldoRecordAcademico(
                                        recordacademicooriginal=recordantiguo,
                                        recordacademiconuevo=recordnuevo
                                    )
                                    respaldorecord.save()
                                else:
                                    respaldorecord = respaldo[0]
                                    respaldorecord.recordacademiconuevo = recordnuevo
                                    respaldorecord.save()
                                print(u"Record actualizado %s" % recordnuevo)


                        else:
                            hojadestino.write(fila, 3, "NO ENCONTRO RECORD ANTIGUO %s" % equivalencia.asignaturamalla)
                            fila += 1

                    fila += 1

                    time.sleep(3)

            lin += 1

        libdestino.save(output_folder + libre_origen)
        print(output_folder + libre_origen)
        print("Proceso finalizado. . .")


    except Exception as ex:
        transaction.set_rollback(True)
        print('error: %s' % ex)
        hojadestino.write(fila, 3, str(ex))
        fila += 1
def homologacion_derecho_2():
    # verificar que todos los estudiantes tengan la misma malla
    cadena = ''
    linea, excluidos, conexito = 0, 0, 0
    try:
        libre_origen = '/homologacion_derecho_2_2.xls'
        fuentecabecera = easyxf(
            'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
        fuentenormal = easyxf(
            'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
        output_folder = MEDIA_ROOT
        output_folder = os.path.join(os.path.join(BASE_DIR))
        # liborigen = xlrd.open_workbook(output_folder + libre_origen)
        libdestino = xlwt.Workbook()
        hojadestino = libdestino.add_sheet('HOJA1')
        fil = 0
        columnas = [(u"CEDULA", 6000),
                    (u"APELLIDOS Y NOMBRES", 6000),
                    (u"OBSERVACIÓN", 6000)
                    ]
        for col_num in range(len(columnas)):
            hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
            hojadestino.col(col_num).width = columnas[col_num][1]
        fila = 1
        titulacion = 0

        lin = 0
        miarchivo = openpyxl.load_workbook("derecholineaf.xlsx")
        # miarchivo = openpyxl.load_workbook("prueba_2.xlsx")

        ws = miarchivo.get_sheet_by_name("segundo")
        worksheet = ws
        c = 0
        cont = 0
        periodo_id = 177
        carrera_id = 126
        mallaantigua_id = 207
        mallanueva_id = 480

        for row in worksheet.iter_rows(min_row=0):
            if lin >= 0:
                currentValues, cadena = [], ''
                for cell in row:
                    cadena += str(cell.value) + ' '
                    currentValues.append(str(cell.value))
                identificacion = currentValues[0]

                with transaction.atomic():
                    matricula = Matricula.objects.filter(status=True, inscripcion__carrera__id=carrera_id,
                                                         inscripcion__persona__cedula=identificacion).first()
                    cont += 1
                    matricula.pasoayuda = True
                    matricula.save()
                    print(u"%s - %s" % (matricula, cont))
                    inscripcion = matricula.inscripcion
                    hojadestino.write(fila, 0, matricula.inscripcion.persona.identificacion(), fuentenormal)
                    hojadestino.write(fila, 1, matricula.inscripcion.persona.nombre_completo(), fuentenormal)
                    hojadestino.write(fila, 2, matricula.nivel.periodo.nombre, fuentenormal)

                    if InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                       malla_id=mallaantigua_id).exists():
                        imantigua = \
                        InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallaantigua_id)[
                            0]
                        imantigua.status = False
                        imantigua.save()
                        print(u"Desactiva antigua inscripcion -----------------------------")

                    if not InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                           malla_id=mallanueva_id).exists():
                        imnueva = InscripcionMalla(inscripcion=inscripcion, malla_id=mallanueva_id)
                        imnueva.save()
                        print(u"Crea nueva inscripcion -----------------------------")

                    equivalencias = TablaEquivalenciaAsignaturas.objects.filter(status=True,
                                                                                asignaturamalla__malla_id=mallaantigua_id).order_by(
                        'asignaturamallasalto__nivelmalla__orden')
                    for equivalencia in equivalencias:
                        print(u"nueva - %s" % equivalencia.asignaturamallasalto)
                        recordantiguo = inscripcion.recordacademico_set.filter(status=True,
                                                                               asignaturamalla_id=equivalencia.asignaturamalla_id).first()

                        if recordantiguo:
                            print(u"anterior - %s" % equivalencia.asignaturamalla)
                            print(u"Record antiguo: %s" % recordantiguo)
                            recordnuevo = None
                            recordantiguo.status = False
                            recordantiguo.save(update_asignaturamalla=False)

                            if equivalencia.asignaturamallasalto_id == 10623 or equivalencia.asignaturamallasalto_id == 10627 or equivalencia.asignaturamallasalto_id == 10615:
                                observaciones = recordantiguo.observaciones + " Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15"
                                homologada = True
                            else:
                                observaciones = recordantiguo.observaciones + " Migración con base a RESOLUCIÓN CGA-SO-9-2023-NO15 Y RESOLUCIÓN OCS-SO-18-2023-NO14"
                                homologada = False
                            if not RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                  asignaturamalla=equivalencia.asignaturamallasalto).exists():

                                recordnuevo = RecordAcademico(inscripcion=inscripcion,
                                                              matriculas=recordantiguo.matriculas,
                                                              asignaturamalla=equivalencia.asignaturamallasalto,
                                                              asignatura=equivalencia.asignaturamallasalto.asignatura,
                                                              asignaturaold_id=recordantiguo.asignatura.id,
                                                              nota=recordantiguo.nota,
                                                              asistencia=recordantiguo.asistencia,
                                                              sinasistencia=recordantiguo.sinasistencia,
                                                              fecha=recordantiguo.fecha,
                                                              noaplica=recordantiguo.noaplica,
                                                              aprobada=recordantiguo.aprobada,
                                                              convalidacion=recordantiguo.convalidacion,
                                                              pendiente=recordantiguo.pendiente,
                                                              creditos=equivalencia.asignaturamallasalto.creditos,
                                                              horas=equivalencia.asignaturamallasalto.horas,
                                                              valida=recordantiguo.valida,
                                                              validapromedio=recordantiguo.validapromedio,
                                                              observaciones=observaciones,
                                                              homologada=homologada,
                                                              materiaregular=recordantiguo.materiaregular,
                                                              materiacurso=None,
                                                              completonota=recordantiguo.completonota,
                                                              completoasistencia=recordantiguo.completoasistencia,
                                                              fechainicio=recordantiguo.fechainicio,
                                                              fechafin=recordantiguo.fechafin,
                                                              suficiencia=recordantiguo.suficiencia,
                                                              asignaturamallahistorico_id=recordantiguo.asignaturamalla.id,
                                                              reverso=False)
                                recordnuevo.save()
                                print(u"Crea nuevo record %s" % recordnuevo)


                            elif RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                asignaturamalla=equivalencia.asignaturamallasalto):
                                recordnuevo = RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                             asignaturamalla=equivalencia.asignaturamallasalto)[
                                    0]
                                recordnuevo.matriculas = recordantiguo.matriculas
                                recordnuevo.asignaturamalla = equivalencia.asignaturamallasalto
                                recordnuevo.asignatura = equivalencia.asignaturamallasalto.asignatura
                                recordnuevo.asignaturaold = recordantiguo.asignaturamalla.asignatura
                                recordnuevo.nota = recordantiguo.nota
                                recordnuevo.asistencia = recordantiguo.asistencia
                                recordnuevo.sinasistencia = recordantiguo.sinasistencia
                                recordnuevo.fecha = recordantiguo.fecha
                                recordnuevo.noaplica = recordantiguo.noaplica
                                recordnuevo.aprobada = recordantiguo.aprobada
                                recordnuevo.convalidacion = recordantiguo.convalidacion
                                recordnuevo.pendiente = recordantiguo.pendiente
                                recordnuevo.creditos = equivalencia.asignaturamallasalto.creditos
                                recordnuevo.horas = equivalencia.asignaturamallasalto.horas
                                recordnuevo.valida = recordantiguo.valida
                                recordnuevo.validapromedio = recordantiguo.validapromedio
                                recordnuevo.observaciones = observaciones
                                recordnuevo.homologada = homologada
                                recordnuevo.materiaregular = recordantiguo.materiaregular
                                recordnuevo.materiacurso = None
                                recordnuevo.completonota = recordantiguo.completonota
                                recordnuevo.completoasistencia = recordantiguo.completoasistencia
                                recordnuevo.fechainicio = recordantiguo.fechainicio
                                recordnuevo.fechafin = recordantiguo.fechafin
                                recordnuevo.suficiencia = recordantiguo.suficiencia
                                recordnuevo.asignaturamallahistorico = recordantiguo.asignaturamalla
                                recordnuevo.reverso = False
                                recordnuevo.save()

                            if recordnuevo:
                                historicos = HistoricoRecordAcademico.objects.filter(status=True,
                                                                                     recordacademico=recordantiguo).update(
                                    recordacademico=recordnuevo,
                                    creditos=recordnuevo.creditos,
                                    horas=recordnuevo.horas,
                                    homologada=recordnuevo.homologada)
                                respaldo = RespaldoRecordAcademico.objects.filter(status=True,
                                                                                  recordacademicooriginal=recordantiguo)

                                if not respaldo.exists():
                                    respaldorecord = RespaldoRecordAcademico(
                                        recordacademicooriginal=recordantiguo,
                                        recordacademiconuevo=recordnuevo
                                    )
                                    respaldorecord.save()
                                else:
                                    respaldorecord = respaldo[0]
                                    respaldorecord.recordacademiconuevo = recordnuevo
                                    respaldorecord.save()
                                print(u"Record actualizado %s" % recordnuevo)


                        else:
                            hojadestino.write(fila, 3, "NO ENCONTRO RECORD ANTIGUO %s" % equivalencia.asignaturamalla)
                            fila += 1

                    fila += 1

                    time.sleep(3)

            lin += 1

        libdestino.save(output_folder + libre_origen)
        print(output_folder + libre_origen)
        print("Proceso finalizado. . .")


    except Exception as ex:
        transaction.set_rollback(True)
        print('error: %s' % ex)
        hojadestino.write(fila, 3, str(ex))
        fila += 1
def homologacion_derecho_3():
    # verificar que todos los estudiantes tengan la misma malla
    cadena = ''
    linea, excluidos, conexito = 0, 0, 0
    try:
        libre_origen = '/homologacion_derecho_2_3.xls'
        fuentecabecera = easyxf(
            'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
        fuentenormal = easyxf(
            'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
        output_folder = MEDIA_ROOT
        output_folder = os.path.join(os.path.join(BASE_DIR))
        # liborigen = xlrd.open_workbook(output_folder + libre_origen)
        libdestino = xlwt.Workbook()
        hojadestino = libdestino.add_sheet('HOJA1')
        fil = 0
        columnas = [(u"CEDULA", 6000),
                    (u"APELLIDOS Y NOMBRES", 6000),
                    (u"OBSERVACIÓN", 6000)
                    ]
        for col_num in range(len(columnas)):
            hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
            hojadestino.col(col_num).width = columnas[col_num][1]
        fila = 1
        titulacion = 0

        lin = 0
        miarchivo = openpyxl.load_workbook("derecholineaf.xlsx")
        # miarchivo = openpyxl.load_workbook("prueba_2.xlsx")

        ws = miarchivo.get_sheet_by_name("tercero")
        worksheet = ws
        c = 0
        cont = 0
        periodo_id = 177
        carrera_id = 126
        mallaantigua_id = 207
        mallanueva_id = 480

        for row in worksheet.iter_rows(min_row=0):
            if lin >= 0:
                currentValues, cadena = [], ''
                for cell in row:
                    cadena += str(cell.value) + ' '
                    currentValues.append(str(cell.value))
                identificacion = currentValues[0]

                with transaction.atomic():
                    matricula = Matricula.objects.filter(status=True, inscripcion__carrera__id=carrera_id,
                                                         inscripcion__persona__cedula=identificacion).first()
                    cont += 1
                    matricula.pasoayuda = True
                    matricula.save()
                    print(u"%s - %s" % (matricula, cont))
                    inscripcion = matricula.inscripcion
                    hojadestino.write(fila, 0, matricula.inscripcion.persona.identificacion(), fuentenormal)
                    hojadestino.write(fila, 1, matricula.inscripcion.persona.nombre_completo(), fuentenormal)
                    hojadestino.write(fila, 2, matricula.nivel.periodo.nombre, fuentenormal)

                    if InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                       malla_id=mallaantigua_id).exists():
                        imantigua = \
                        InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallaantigua_id)[
                            0]
                        imantigua.status = False
                        imantigua.save()
                        print(u"Desactiva antigua inscripcion -----------------------------")

                    if not InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                           malla_id=mallanueva_id).exists():
                        imnueva = InscripcionMalla(inscripcion=inscripcion, malla_id=mallanueva_id)
                        imnueva.save()
                        print(u"Crea nueva inscripcion -----------------------------")

                    equivalencias = TablaEquivalenciaAsignaturas.objects.filter(status=True,
                                                                                asignaturamalla__malla_id=mallaantigua_id).order_by(
                        'asignaturamallasalto__nivelmalla__orden')
                    for equivalencia in equivalencias:
                        print(u"nueva - %s" % equivalencia.asignaturamallasalto)
                        recordantiguo = inscripcion.recordacademico_set.filter(status=True,
                                                                               asignaturamalla_id=equivalencia.asignaturamalla_id).first()

                        if recordantiguo:
                            print(u"anterior - %s" % equivalencia.asignaturamalla)
                            print(u"Record antiguo: %s" % recordantiguo)
                            recordnuevo = None
                            recordantiguo.status = False
                            recordantiguo.save(update_asignaturamalla=False)

                            if equivalencia.asignaturamallasalto_id == 10623 or equivalencia.asignaturamallasalto_id == 10627 or equivalencia.asignaturamallasalto_id == 10615:
                                observaciones = recordantiguo.observaciones + " Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15"
                                homologada = True
                            else:
                                observaciones = recordantiguo.observaciones + " Migración con base a RESOLUCIÓN CGA-SO-9-2023-NO15 Y RESOLUCIÓN OCS-SO-18-2023-NO14"
                                homologada = False
                            if not RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                  asignaturamalla=equivalencia.asignaturamallasalto).exists():

                                recordnuevo = RecordAcademico(inscripcion=inscripcion,
                                                              matriculas=recordantiguo.matriculas,
                                                              asignaturamalla=equivalencia.asignaturamallasalto,
                                                              asignatura=equivalencia.asignaturamallasalto.asignatura,
                                                              asignaturaold_id=recordantiguo.asignatura.id,
                                                              nota=recordantiguo.nota,
                                                              asistencia=recordantiguo.asistencia,
                                                              sinasistencia=recordantiguo.sinasistencia,
                                                              fecha=recordantiguo.fecha,
                                                              noaplica=recordantiguo.noaplica,
                                                              aprobada=recordantiguo.aprobada,
                                                              convalidacion=recordantiguo.convalidacion,
                                                              pendiente=recordantiguo.pendiente,
                                                              creditos=equivalencia.asignaturamallasalto.creditos,
                                                              horas=equivalencia.asignaturamallasalto.horas,
                                                              valida=recordantiguo.valida,
                                                              validapromedio=recordantiguo.validapromedio,
                                                              observaciones=observaciones,
                                                              homologada=homologada,
                                                              materiaregular=recordantiguo.materiaregular,
                                                              materiacurso=None,
                                                              completonota=recordantiguo.completonota,
                                                              completoasistencia=recordantiguo.completoasistencia,
                                                              fechainicio=recordantiguo.fechainicio,
                                                              fechafin=recordantiguo.fechafin,
                                                              suficiencia=recordantiguo.suficiencia,
                                                              asignaturamallahistorico_id=recordantiguo.asignaturamalla.id,
                                                              reverso=False)
                                recordnuevo.save()
                                print(u"Crea nuevo record %s" % recordnuevo)


                            elif RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                asignaturamalla=equivalencia.asignaturamallasalto):
                                recordnuevo = RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                             asignaturamalla=equivalencia.asignaturamallasalto)[
                                    0]
                                recordnuevo.matriculas = recordantiguo.matriculas
                                recordnuevo.asignaturamalla = equivalencia.asignaturamallasalto
                                recordnuevo.asignatura = equivalencia.asignaturamallasalto.asignatura
                                recordnuevo.asignaturaold = recordantiguo.asignaturamalla.asignatura
                                recordnuevo.nota = recordantiguo.nota
                                recordnuevo.asistencia = recordantiguo.asistencia
                                recordnuevo.sinasistencia = recordantiguo.sinasistencia
                                recordnuevo.fecha = recordantiguo.fecha
                                recordnuevo.noaplica = recordantiguo.noaplica
                                recordnuevo.aprobada = recordantiguo.aprobada
                                recordnuevo.convalidacion = recordantiguo.convalidacion
                                recordnuevo.pendiente = recordantiguo.pendiente
                                recordnuevo.creditos = equivalencia.asignaturamallasalto.creditos
                                recordnuevo.horas = equivalencia.asignaturamallasalto.horas
                                recordnuevo.valida = recordantiguo.valida
                                recordnuevo.validapromedio = recordantiguo.validapromedio
                                recordnuevo.observaciones = observaciones
                                recordnuevo.homologada = homologada
                                recordnuevo.materiaregular = recordantiguo.materiaregular
                                recordnuevo.materiacurso = None
                                recordnuevo.completonota = recordantiguo.completonota
                                recordnuevo.completoasistencia = recordantiguo.completoasistencia
                                recordnuevo.fechainicio = recordantiguo.fechainicio
                                recordnuevo.fechafin = recordantiguo.fechafin
                                recordnuevo.suficiencia = recordantiguo.suficiencia
                                recordnuevo.asignaturamallahistorico = recordantiguo.asignaturamalla
                                recordnuevo.reverso = False
                                recordnuevo.save()

                            if recordnuevo:
                                historicos = HistoricoRecordAcademico.objects.filter(status=True,
                                                                                     recordacademico=recordantiguo).update(
                                    recordacademico=recordnuevo,
                                    creditos=recordnuevo.creditos,
                                    horas=recordnuevo.horas,
                                    homologada=recordnuevo.homologada)
                                respaldo = RespaldoRecordAcademico.objects.filter(status=True,
                                                                                  recordacademicooriginal=recordantiguo)

                                if not respaldo.exists():
                                    respaldorecord = RespaldoRecordAcademico(
                                        recordacademicooriginal=recordantiguo,
                                        recordacademiconuevo=recordnuevo
                                    )
                                    respaldorecord.save()
                                else:
                                    respaldorecord = respaldo[0]
                                    respaldorecord.recordacademiconuevo = recordnuevo
                                    respaldorecord.save()
                                print(u"Record actualizado %s" % recordnuevo)


                        else:
                            hojadestino.write(fila, 3, "NO ENCONTRO RECORD ANTIGUO %s" % equivalencia.asignaturamalla)
                            fila += 1

                    fila += 1

                    time.sleep(3)

            lin += 1

        libdestino.save(output_folder + libre_origen)
        print(output_folder + libre_origen)
        print("Proceso finalizado. . .")


    except Exception as ex:
        transaction.set_rollback(True)
        print('error: %s' % ex)
        hojadestino.write(fila, 3, str(ex))
        fila += 1
def homologacion_derecho_4():
    # verificar que todos los estudiantes tengan la misma malla
    cadena = ''
    linea, excluidos, conexito = 0, 0, 0
    try:
        libre_origen = '/homologacion_derecho_2_4.xls'
        fuentecabecera = easyxf(
            'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
        fuentenormal = easyxf(
            'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
        output_folder = MEDIA_ROOT
        output_folder = os.path.join(os.path.join(BASE_DIR))
        # liborigen = xlrd.open_workbook(output_folder + libre_origen)
        libdestino = xlwt.Workbook()
        hojadestino = libdestino.add_sheet('HOJA1')
        fil = 0
        columnas = [(u"CEDULA", 6000),
                    (u"APELLIDOS Y NOMBRES", 6000),
                    (u"OBSERVACIÓN", 6000)
                    ]
        for col_num in range(len(columnas)):
            hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
            hojadestino.col(col_num).width = columnas[col_num][1]
        fila = 1
        titulacion = 0

        lin = 0
        miarchivo = openpyxl.load_workbook("derecholineaf.xlsx")
        # miarchivo = openpyxl.load_workbook("prueba_2.xlsx")

        ws = miarchivo.get_sheet_by_name("cuarto")
        worksheet = ws
        c = 0
        cont = 0
        periodo_id = 177
        carrera_id = 126
        mallaantigua_id = 207
        mallanueva_id = 480

        for row in worksheet.iter_rows(min_row=0):
            if lin >= 0:
                currentValues, cadena = [], ''
                for cell in row:
                    cadena += str(cell.value) + ' '
                    currentValues.append(str(cell.value))
                identificacion = currentValues[0]

                with transaction.atomic():
                    matricula = Matricula.objects.filter(status=True, inscripcion__carrera__id=carrera_id,
                                                         inscripcion__persona__cedula=identificacion).first()
                    cont += 1
                    matricula.pasoayuda = True
                    matricula.save()
                    print(u"%s - %s" % (matricula, cont))
                    inscripcion = matricula.inscripcion
                    hojadestino.write(fila, 0, matricula.inscripcion.persona.identificacion(), fuentenormal)
                    hojadestino.write(fila, 1, matricula.inscripcion.persona.nombre_completo(), fuentenormal)
                    hojadestino.write(fila, 2, matricula.nivel.periodo.nombre, fuentenormal)

                    if InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                       malla_id=mallaantigua_id).exists():
                        imantigua = \
                        InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallaantigua_id)[
                            0]
                        imantigua.status = False
                        imantigua.save()
                        print(u"Desactiva antigua inscripcion -----------------------------")

                    if not InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                           malla_id=mallanueva_id).exists():
                        imnueva = InscripcionMalla(inscripcion=inscripcion, malla_id=mallanueva_id)
                        imnueva.save()
                        print(u"Crea nueva inscripcion -----------------------------")

                    equivalencias = TablaEquivalenciaAsignaturas.objects.filter(status=True,
                                                                                asignaturamalla__malla_id=mallaantigua_id).order_by(
                        'asignaturamallasalto__nivelmalla__orden')
                    for equivalencia in equivalencias:
                        print(u"nueva - %s" % equivalencia.asignaturamallasalto)
                        recordantiguo = inscripcion.recordacademico_set.filter(status=True,
                                                                               asignaturamalla_id=equivalencia.asignaturamalla_id).first()

                        if recordantiguo:
                            print(u"anterior - %s" % equivalencia.asignaturamalla)
                            print(u"Record antiguo: %s" % recordantiguo)
                            recordnuevo = None
                            recordantiguo.status = False
                            recordantiguo.save(update_asignaturamalla=False)

                            if equivalencia.asignaturamallasalto_id == 10623 or equivalencia.asignaturamallasalto_id == 10627 or equivalencia.asignaturamallasalto_id == 10615:
                                observaciones = recordantiguo.observaciones + " Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15"
                                homologada = True
                            else:
                                observaciones = recordantiguo.observaciones + " Migración con base a RESOLUCIÓN CGA-SO-9-2023-NO15 Y RESOLUCIÓN OCS-SO-18-2023-NO14"
                                homologada = False
                            if not RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                  asignaturamalla=equivalencia.asignaturamallasalto).exists():

                                recordnuevo = RecordAcademico(inscripcion=inscripcion,
                                                              matriculas=recordantiguo.matriculas,
                                                              asignaturamalla=equivalencia.asignaturamallasalto,
                                                              asignatura=equivalencia.asignaturamallasalto.asignatura,
                                                              asignaturaold_id=recordantiguo.asignatura.id,
                                                              nota=recordantiguo.nota,
                                                              asistencia=recordantiguo.asistencia,
                                                              sinasistencia=recordantiguo.sinasistencia,
                                                              fecha=recordantiguo.fecha,
                                                              noaplica=recordantiguo.noaplica,
                                                              aprobada=recordantiguo.aprobada,
                                                              convalidacion=recordantiguo.convalidacion,
                                                              pendiente=recordantiguo.pendiente,
                                                              creditos=equivalencia.asignaturamallasalto.creditos,
                                                              horas=equivalencia.asignaturamallasalto.horas,
                                                              valida=recordantiguo.valida,
                                                              validapromedio=recordantiguo.validapromedio,
                                                              observaciones=observaciones,
                                                              homologada=homologada,
                                                              materiaregular=recordantiguo.materiaregular,
                                                              materiacurso=None,
                                                              completonota=recordantiguo.completonota,
                                                              completoasistencia=recordantiguo.completoasistencia,
                                                              fechainicio=recordantiguo.fechainicio,
                                                              fechafin=recordantiguo.fechafin,
                                                              suficiencia=recordantiguo.suficiencia,
                                                              asignaturamallahistorico_id=recordantiguo.asignaturamalla.id,
                                                              reverso=False)
                                recordnuevo.save()
                                print(u"Crea nuevo record %s" % recordnuevo)


                            elif RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                asignaturamalla=equivalencia.asignaturamallasalto):
                                recordnuevo = RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                             asignaturamalla=equivalencia.asignaturamallasalto)[
                                    0]
                                recordnuevo.matriculas = recordantiguo.matriculas
                                recordnuevo.asignaturamalla = equivalencia.asignaturamallasalto
                                recordnuevo.asignatura = equivalencia.asignaturamallasalto.asignatura
                                recordnuevo.asignaturaold = recordantiguo.asignaturamalla.asignatura
                                recordnuevo.nota = recordantiguo.nota
                                recordnuevo.asistencia = recordantiguo.asistencia
                                recordnuevo.sinasistencia = recordantiguo.sinasistencia
                                recordnuevo.fecha = recordantiguo.fecha
                                recordnuevo.noaplica = recordantiguo.noaplica
                                recordnuevo.aprobada = recordantiguo.aprobada
                                recordnuevo.convalidacion = recordantiguo.convalidacion
                                recordnuevo.pendiente = recordantiguo.pendiente
                                recordnuevo.creditos = equivalencia.asignaturamallasalto.creditos
                                recordnuevo.horas = equivalencia.asignaturamallasalto.horas
                                recordnuevo.valida = recordantiguo.valida
                                recordnuevo.validapromedio = recordantiguo.validapromedio
                                recordnuevo.observaciones = observaciones
                                recordnuevo.homologada = homologada
                                recordnuevo.materiaregular = recordantiguo.materiaregular
                                recordnuevo.materiacurso = None
                                recordnuevo.completonota = recordantiguo.completonota
                                recordnuevo.completoasistencia = recordantiguo.completoasistencia
                                recordnuevo.fechainicio = recordantiguo.fechainicio
                                recordnuevo.fechafin = recordantiguo.fechafin
                                recordnuevo.suficiencia = recordantiguo.suficiencia
                                recordnuevo.asignaturamallahistorico = recordantiguo.asignaturamalla
                                recordnuevo.reverso = False
                                recordnuevo.save()

                            if recordnuevo:
                                historicos = HistoricoRecordAcademico.objects.filter(status=True,
                                                                                     recordacademico=recordantiguo).update(
                                    recordacademico=recordnuevo,
                                    creditos=recordnuevo.creditos,
                                    horas=recordnuevo.horas,
                                    homologada=recordnuevo.homologada)
                                respaldo = RespaldoRecordAcademico.objects.filter(status=True,
                                                                                  recordacademicooriginal=recordantiguo)

                                if not respaldo.exists():
                                    respaldorecord = RespaldoRecordAcademico(
                                        recordacademicooriginal=recordantiguo,
                                        recordacademiconuevo=recordnuevo
                                    )
                                    respaldorecord.save()
                                else:
                                    respaldorecord = respaldo[0]
                                    respaldorecord.recordacademiconuevo = recordnuevo
                                    respaldorecord.save()
                                print(u"Record actualizado %s" % recordnuevo)


                        else:
                            hojadestino.write(fila, 3, "NO ENCONTRO RECORD ANTIGUO %s" % equivalencia.asignaturamalla)
                            fila += 1

                    fila += 1

                    time.sleep(3)

            lin += 1

        libdestino.save(output_folder + libre_origen)
        print(output_folder + libre_origen)
        print("Proceso finalizado. . .")


    except Exception as ex:
        transaction.set_rollback(True)
        print('error: %s' % ex)
        hojadestino.write(fila, 3, str(ex))
        fila += 1
def homologacion_derecho_5():
    # verificar que todos los estudiantes tengan la misma malla
    cadena = ''
    linea, excluidos, conexito = 0, 0, 0
    try:
        libre_origen = '/homologacion_derecho_2_5.xls'
        fuentecabecera = easyxf(
            'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
        fuentenormal = easyxf(
            'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
        output_folder = MEDIA_ROOT
        output_folder = os.path.join(os.path.join(BASE_DIR))
        # liborigen = xlrd.open_workbook(output_folder + libre_origen)
        libdestino = xlwt.Workbook()
        hojadestino = libdestino.add_sheet('HOJA1')
        fil = 0
        columnas = [(u"CEDULA", 6000),
                    (u"APELLIDOS Y NOMBRES", 6000),
                    (u"OBSERVACIÓN", 6000)
                    ]
        for col_num in range(len(columnas)):
            hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
            hojadestino.col(col_num).width = columnas[col_num][1]
        fila = 1
        titulacion = 0

        lin = 0
        miarchivo = openpyxl.load_workbook("derecholineaf.xlsx")
        # miarchivo = openpyxl.load_workbook("prueba_2.xlsx")

        ws = miarchivo.get_sheet_by_name("quinto")
        worksheet = ws
        c = 0
        cont = 0
        periodo_id = 177
        carrera_id = 126
        mallaantigua_id = 207
        mallanueva_id = 480

        for row in worksheet.iter_rows(min_row=0):
            if lin >= 0:
                currentValues, cadena = [], ''
                for cell in row:
                    cadena += str(cell.value) + ' '
                    currentValues.append(str(cell.value))
                identificacion = currentValues[0]

                with transaction.atomic():
                    matricula = Matricula.objects.filter(status=True, inscripcion__carrera__id=carrera_id,
                                                         inscripcion__persona__cedula=identificacion).first()
                    cont += 1
                    matricula.pasoayuda = True
                    matricula.save()
                    print(u"%s - %s" % (matricula, cont))
                    inscripcion = matricula.inscripcion
                    hojadestino.write(fila, 0, matricula.inscripcion.persona.identificacion(), fuentenormal)
                    hojadestino.write(fila, 1, matricula.inscripcion.persona.nombre_completo(), fuentenormal)
                    hojadestino.write(fila, 2, matricula.nivel.periodo.nombre, fuentenormal)

                    if InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                       malla_id=mallaantigua_id).exists():
                        imantigua = \
                        InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallaantigua_id)[
                            0]
                        imantigua.status = False
                        imantigua.save()
                        print(u"Desactiva antigua inscripcion -----------------------------")

                    if not InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                           malla_id=mallanueva_id).exists():
                        imnueva = InscripcionMalla(inscripcion=inscripcion, malla_id=mallanueva_id)
                        imnueva.save()
                        print(u"Crea nueva inscripcion -----------------------------")

                    equivalencias = TablaEquivalenciaAsignaturas.objects.filter(status=True,
                                                                                asignaturamalla__malla_id=mallaantigua_id).order_by(
                        'asignaturamallasalto__nivelmalla__orden')
                    for equivalencia in equivalencias:
                        print(u"nueva - %s" % equivalencia.asignaturamallasalto)
                        recordantiguo = inscripcion.recordacademico_set.filter(status=True,
                                                                               asignaturamalla_id=equivalencia.asignaturamalla_id).first()

                        if recordantiguo:
                            print(u"anterior - %s" % equivalencia.asignaturamalla)
                            print(u"Record antiguo: %s" % recordantiguo)
                            recordnuevo = None
                            recordantiguo.status = False
                            recordantiguo.save(update_asignaturamalla=False)

                            if equivalencia.asignaturamallasalto_id == 10623 or equivalencia.asignaturamallasalto_id == 10627 or equivalencia.asignaturamallasalto_id == 10615:
                                observaciones = recordantiguo.observaciones + " Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15"
                                homologada = True
                            else:
                                observaciones = recordantiguo.observaciones + " Migración con base a RESOLUCIÓN CGA-SO-9-2023-NO15 Y RESOLUCIÓN OCS-SO-18-2023-NO14"
                                homologada = False
                            if not RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                  asignaturamalla=equivalencia.asignaturamallasalto).exists():

                                recordnuevo = RecordAcademico(inscripcion=inscripcion,
                                                              matriculas=recordantiguo.matriculas,
                                                              asignaturamalla=equivalencia.asignaturamallasalto,
                                                              asignatura=equivalencia.asignaturamallasalto.asignatura,
                                                              asignaturaold_id=recordantiguo.asignatura.id,
                                                              nota=recordantiguo.nota,
                                                              asistencia=recordantiguo.asistencia,
                                                              sinasistencia=recordantiguo.sinasistencia,
                                                              fecha=recordantiguo.fecha,
                                                              noaplica=recordantiguo.noaplica,
                                                              aprobada=recordantiguo.aprobada,
                                                              convalidacion=recordantiguo.convalidacion,
                                                              pendiente=recordantiguo.pendiente,
                                                              creditos=equivalencia.asignaturamallasalto.creditos,
                                                              horas=equivalencia.asignaturamallasalto.horas,
                                                              valida=recordantiguo.valida,
                                                              validapromedio=recordantiguo.validapromedio,
                                                              observaciones=observaciones,
                                                              homologada=homologada,
                                                              materiaregular=recordantiguo.materiaregular,
                                                              materiacurso=None,
                                                              completonota=recordantiguo.completonota,
                                                              completoasistencia=recordantiguo.completoasistencia,
                                                              fechainicio=recordantiguo.fechainicio,
                                                              fechafin=recordantiguo.fechafin,
                                                              suficiencia=recordantiguo.suficiencia,
                                                              asignaturamallahistorico_id=recordantiguo.asignaturamalla.id,
                                                              reverso=False)
                                recordnuevo.save()
                                print(u"Crea nuevo record %s" % recordnuevo)


                            elif RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                asignaturamalla=equivalencia.asignaturamallasalto):
                                recordnuevo = RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                             asignaturamalla=equivalencia.asignaturamallasalto)[
                                    0]
                                recordnuevo.matriculas = recordantiguo.matriculas
                                recordnuevo.asignaturamalla = equivalencia.asignaturamallasalto
                                recordnuevo.asignatura = equivalencia.asignaturamallasalto.asignatura
                                recordnuevo.asignaturaold = recordantiguo.asignaturamalla.asignatura
                                recordnuevo.nota = recordantiguo.nota
                                recordnuevo.asistencia = recordantiguo.asistencia
                                recordnuevo.sinasistencia = recordantiguo.sinasistencia
                                recordnuevo.fecha = recordantiguo.fecha
                                recordnuevo.noaplica = recordantiguo.noaplica
                                recordnuevo.aprobada = recordantiguo.aprobada
                                recordnuevo.convalidacion = recordantiguo.convalidacion
                                recordnuevo.pendiente = recordantiguo.pendiente
                                recordnuevo.creditos = equivalencia.asignaturamallasalto.creditos
                                recordnuevo.horas = equivalencia.asignaturamallasalto.horas
                                recordnuevo.valida = recordantiguo.valida
                                recordnuevo.validapromedio = recordantiguo.validapromedio
                                recordnuevo.observaciones = observaciones
                                recordnuevo.homologada = homologada
                                recordnuevo.materiaregular = recordantiguo.materiaregular
                                recordnuevo.materiacurso = None
                                recordnuevo.completonota = recordantiguo.completonota
                                recordnuevo.completoasistencia = recordantiguo.completoasistencia
                                recordnuevo.fechainicio = recordantiguo.fechainicio
                                recordnuevo.fechafin = recordantiguo.fechafin
                                recordnuevo.suficiencia = recordantiguo.suficiencia
                                recordnuevo.asignaturamallahistorico = recordantiguo.asignaturamalla
                                recordnuevo.reverso = False
                                recordnuevo.save()

                            if recordnuevo:
                                historicos = HistoricoRecordAcademico.objects.filter(status=True,
                                                                                     recordacademico=recordantiguo).update(
                                    recordacademico=recordnuevo,
                                    creditos=recordnuevo.creditos,
                                    horas=recordnuevo.horas,
                                    homologada=recordnuevo.homologada)
                                respaldo = RespaldoRecordAcademico.objects.filter(status=True,
                                                                                  recordacademicooriginal=recordantiguo)

                                if not respaldo.exists():
                                    respaldorecord = RespaldoRecordAcademico(
                                        recordacademicooriginal=recordantiguo,
                                        recordacademiconuevo=recordnuevo
                                    )
                                    respaldorecord.save()
                                else:
                                    respaldorecord = respaldo[0]
                                    respaldorecord.recordacademiconuevo = recordnuevo
                                    respaldorecord.save()
                                print(u"Record actualizado %s" % recordnuevo)


                        else:
                            hojadestino.write(fila, 3, "NO ENCONTRO RECORD ANTIGUO %s" % equivalencia.asignaturamalla)
                            fila += 1

                    fila += 1

                    time.sleep(3)

            lin += 1

        libdestino.save(output_folder + libre_origen)
        print(output_folder + libre_origen)
        print("Proceso finalizado. . .")


    except Exception as ex:
        transaction.set_rollback(True)
        print('error: %s' % ex)
        hojadestino.write(fila, 3, str(ex))
        fila += 1
def homologacion_derecho_6():
    # verificar que todos los estudiantes tengan la misma malla
    cadena = ''
    linea, excluidos, conexito = 0, 0, 0
    try:
        libre_origen = '/homologacion_derecho_2_6.xls'
        fuentecabecera = easyxf(
            'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
        fuentenormal = easyxf(
            'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
        output_folder = MEDIA_ROOT
        output_folder = os.path.join(os.path.join(BASE_DIR))
        # liborigen = xlrd.open_workbook(output_folder + libre_origen)
        libdestino = xlwt.Workbook()
        hojadestino = libdestino.add_sheet('HOJA1')
        fil = 0
        columnas = [(u"CEDULA", 6000),
                    (u"APELLIDOS Y NOMBRES", 6000),
                    (u"OBSERVACIÓN", 6000)
                    ]
        for col_num in range(len(columnas)):
            hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
            hojadestino.col(col_num).width = columnas[col_num][1]
        fila = 1
        titulacion = 0

        lin = 0
        miarchivo = openpyxl.load_workbook("derecholineaf.xlsx")
        # miarchivo = openpyxl.load_workbook("prueba_2.xlsx")

        ws = miarchivo.get_sheet_by_name("sexto")
        worksheet = ws
        c = 0
        cont = 0
        periodo_id = 177
        carrera_id = 126
        mallaantigua_id = 207
        mallanueva_id = 480

        for row in worksheet.iter_rows(min_row=0):
            if lin >= 0:
                currentValues, cadena = [], ''
                for cell in row:
                    cadena += str(cell.value) + ' '
                    currentValues.append(str(cell.value))
                identificacion = currentValues[0]

                with transaction.atomic():
                    matricula = Matricula.objects.filter(status=True, inscripcion__carrera__id=carrera_id,
                                                         inscripcion__persona__cedula=identificacion).first()
                    cont += 1
                    matricula.pasoayuda = True
                    matricula.save()
                    print(u"%s - %s" % (matricula, cont))
                    inscripcion = matricula.inscripcion
                    hojadestino.write(fila, 0, matricula.inscripcion.persona.identificacion(), fuentenormal)
                    hojadestino.write(fila, 1, matricula.inscripcion.persona.nombre_completo(), fuentenormal)
                    hojadestino.write(fila, 2, matricula.nivel.periodo.nombre, fuentenormal)

                    if InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                       malla_id=mallaantigua_id).exists():
                        imantigua = \
                        InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallaantigua_id)[
                            0]
                        imantigua.status = False
                        imantigua.save()
                        print(u"Desactiva antigua inscripcion -----------------------------")

                    if not InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                           malla_id=mallanueva_id).exists():
                        imnueva = InscripcionMalla(inscripcion=inscripcion, malla_id=mallanueva_id)
                        imnueva.save()
                        print(u"Crea nueva inscripcion -----------------------------")

                    equivalencias = TablaEquivalenciaAsignaturas.objects.filter(status=True,
                                                                                asignaturamalla__malla_id=mallaantigua_id).order_by(
                        'asignaturamallasalto__nivelmalla__orden')
                    for equivalencia in equivalencias:
                        print(u"nueva - %s" % equivalencia.asignaturamallasalto)
                        recordantiguo = inscripcion.recordacademico_set.filter(status=True,
                                                                               asignaturamalla_id=equivalencia.asignaturamalla_id).first()

                        if recordantiguo:
                            print(u"anterior - %s" % equivalencia.asignaturamalla)
                            print(u"Record antiguo: %s" % recordantiguo)
                            recordnuevo = None
                            recordantiguo.status = False
                            recordantiguo.save(update_asignaturamalla=False)

                            if equivalencia.asignaturamallasalto_id == 10623 or equivalencia.asignaturamallasalto_id == 10627 or equivalencia.asignaturamallasalto_id == 10615:
                                observaciones = recordantiguo.observaciones + " Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15"
                                homologada = True
                            else:
                                observaciones = recordantiguo.observaciones + " Migración con base a RESOLUCIÓN CGA-SO-9-2023-NO15 Y RESOLUCIÓN OCS-SO-18-2023-NO14"
                                homologada = False
                            if not RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                  asignaturamalla=equivalencia.asignaturamallasalto).exists():

                                recordnuevo = RecordAcademico(inscripcion=inscripcion,
                                                              matriculas=recordantiguo.matriculas,
                                                              asignaturamalla=equivalencia.asignaturamallasalto,
                                                              asignatura=equivalencia.asignaturamallasalto.asignatura,
                                                              asignaturaold_id=recordantiguo.asignatura.id,
                                                              nota=recordantiguo.nota,
                                                              asistencia=recordantiguo.asistencia,
                                                              sinasistencia=recordantiguo.sinasistencia,
                                                              fecha=recordantiguo.fecha,
                                                              noaplica=recordantiguo.noaplica,
                                                              aprobada=recordantiguo.aprobada,
                                                              convalidacion=recordantiguo.convalidacion,
                                                              pendiente=recordantiguo.pendiente,
                                                              creditos=equivalencia.asignaturamallasalto.creditos,
                                                              horas=equivalencia.asignaturamallasalto.horas,
                                                              valida=recordantiguo.valida,
                                                              validapromedio=recordantiguo.validapromedio,
                                                              observaciones=observaciones,
                                                              homologada=homologada,
                                                              materiaregular=recordantiguo.materiaregular,
                                                              materiacurso=None,
                                                              completonota=recordantiguo.completonota,
                                                              completoasistencia=recordantiguo.completoasistencia,
                                                              fechainicio=recordantiguo.fechainicio,
                                                              fechafin=recordantiguo.fechafin,
                                                              suficiencia=recordantiguo.suficiencia,
                                                              asignaturamallahistorico_id=recordantiguo.asignaturamalla.id,
                                                              reverso=False)
                                recordnuevo.save()
                                print(u"Crea nuevo record %s" % recordnuevo)


                            elif RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                asignaturamalla=equivalencia.asignaturamallasalto):
                                recordnuevo = RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                             asignaturamalla=equivalencia.asignaturamallasalto)[
                                    0]
                                recordnuevo.matriculas = recordantiguo.matriculas
                                recordnuevo.asignaturamalla = equivalencia.asignaturamallasalto
                                recordnuevo.asignatura = equivalencia.asignaturamallasalto.asignatura
                                recordnuevo.asignaturaold = recordantiguo.asignaturamalla.asignatura
                                recordnuevo.nota = recordantiguo.nota
                                recordnuevo.asistencia = recordantiguo.asistencia
                                recordnuevo.sinasistencia = recordantiguo.sinasistencia
                                recordnuevo.fecha = recordantiguo.fecha
                                recordnuevo.noaplica = recordantiguo.noaplica
                                recordnuevo.aprobada = recordantiguo.aprobada
                                recordnuevo.convalidacion = recordantiguo.convalidacion
                                recordnuevo.pendiente = recordantiguo.pendiente
                                recordnuevo.creditos = equivalencia.asignaturamallasalto.creditos
                                recordnuevo.horas = equivalencia.asignaturamallasalto.horas
                                recordnuevo.valida = recordantiguo.valida
                                recordnuevo.validapromedio = recordantiguo.validapromedio
                                recordnuevo.observaciones = observaciones
                                recordnuevo.homologada = homologada
                                recordnuevo.materiaregular = recordantiguo.materiaregular
                                recordnuevo.materiacurso = None
                                recordnuevo.completonota = recordantiguo.completonota
                                recordnuevo.completoasistencia = recordantiguo.completoasistencia
                                recordnuevo.fechainicio = recordantiguo.fechainicio
                                recordnuevo.fechafin = recordantiguo.fechafin
                                recordnuevo.suficiencia = recordantiguo.suficiencia
                                recordnuevo.asignaturamallahistorico = recordantiguo.asignaturamalla
                                recordnuevo.reverso = False
                                recordnuevo.save()

                            if recordnuevo:
                                historicos = HistoricoRecordAcademico.objects.filter(status=True,
                                                                                     recordacademico=recordantiguo).update(
                                    recordacademico=recordnuevo,
                                    creditos=recordnuevo.creditos,
                                    horas=recordnuevo.horas,
                                    homologada=recordnuevo.homologada)
                                respaldo = RespaldoRecordAcademico.objects.filter(status=True,
                                                                                  recordacademicooriginal=recordantiguo)

                                if not respaldo.exists():
                                    respaldorecord = RespaldoRecordAcademico(
                                        recordacademicooriginal=recordantiguo,
                                        recordacademiconuevo=recordnuevo
                                    )
                                    respaldorecord.save()
                                else:
                                    respaldorecord = respaldo[0]
                                    respaldorecord.recordacademiconuevo = recordnuevo
                                    respaldorecord.save()
                                print(u"Record actualizado %s" % recordnuevo)


                        else:
                            hojadestino.write(fila, 3, "NO ENCONTRO RECORD ANTIGUO %s" % equivalencia.asignaturamalla)
                            fila += 1

                    fila += 1

                    time.sleep(3)

            lin += 1

        libdestino.save(output_folder + libre_origen)
        print(output_folder + libre_origen)
        print("Proceso finalizado. . .")


    except Exception as ex:
        transaction.set_rollback(True)
        print('error: %s' % ex)
        hojadestino.write(fila, 3, str(ex))
        fila += 1
def homologacion_derecho_7():
    # verificar que todos los estudiantes tengan la misma malla
    cadena = ''
    linea, excluidos, conexito = 0, 0, 0
    try:
        libre_origen = '/homologacion_derecho_2_7.xls'
        fuentecabecera = easyxf(
            'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
        fuentenormal = easyxf(
            'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
        output_folder = MEDIA_ROOT
        output_folder = os.path.join(os.path.join(BASE_DIR))
        # liborigen = xlrd.open_workbook(output_folder + libre_origen)
        libdestino = xlwt.Workbook()
        hojadestino = libdestino.add_sheet('HOJA1')
        fil = 0
        columnas = [(u"CEDULA", 6000),
                    (u"APELLIDOS Y NOMBRES", 6000),
                    (u"OBSERVACIÓN", 6000)
                    ]
        for col_num in range(len(columnas)):
            hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
            hojadestino.col(col_num).width = columnas[col_num][1]
        fila = 1
        titulacion = 0

        lin = 0
        miarchivo = openpyxl.load_workbook("derecholineaf.xlsx")
        # miarchivo = openpyxl.load_workbook("prueba_2.xlsx")

        ws = miarchivo.get_sheet_by_name("septimo")
        worksheet = ws
        c = 0
        cont = 0
        periodo_id = 177
        carrera_id = 126
        mallaantigua_id = 207
        mallanueva_id = 480

        for row in worksheet.iter_rows(min_row=0):
            if lin >= 0:
                currentValues, cadena = [], ''
                for cell in row:
                    cadena += str(cell.value) + ' '
                    currentValues.append(str(cell.value))
                identificacion = currentValues[0]

                with transaction.atomic():
                    matricula = Matricula.objects.filter(status=True, inscripcion__carrera__id=carrera_id,
                                                         inscripcion__persona__cedula=identificacion).first()
                    cont += 1
                    matricula.pasoayuda = True
                    matricula.save()
                    print(u"%s - %s" % (matricula, cont))
                    inscripcion = matricula.inscripcion
                    hojadestino.write(fila, 0, matricula.inscripcion.persona.identificacion(), fuentenormal)
                    hojadestino.write(fila, 1, matricula.inscripcion.persona.nombre_completo(), fuentenormal)
                    hojadestino.write(fila, 2, matricula.nivel.periodo.nombre, fuentenormal)

                    if InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                       malla_id=mallaantigua_id).exists():
                        imantigua = \
                        InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallaantigua_id)[
                            0]
                        imantigua.status = False
                        imantigua.save()
                        print(u"Desactiva antigua inscripcion -----------------------------")

                    if not InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                           malla_id=mallanueva_id).exists():
                        imnueva = InscripcionMalla(inscripcion=inscripcion, malla_id=mallanueva_id)
                        imnueva.save()
                        print(u"Crea nueva inscripcion -----------------------------")

                    equivalencias = TablaEquivalenciaAsignaturas.objects.filter(status=True,
                                                                                asignaturamalla__malla_id=mallaantigua_id).order_by(
                        'asignaturamallasalto__nivelmalla__orden')
                    for equivalencia in equivalencias:
                        print(u"nueva - %s" % equivalencia.asignaturamallasalto)
                        recordantiguo = inscripcion.recordacademico_set.filter(status=True,
                                                                               asignaturamalla_id=equivalencia.asignaturamalla_id).first()

                        if recordantiguo:
                            print(u"anterior - %s" % equivalencia.asignaturamalla)
                            print(u"Record antiguo: %s" % recordantiguo)
                            recordnuevo = None
                            recordantiguo.status = False
                            recordantiguo.save(update_asignaturamalla=False)

                            if equivalencia.asignaturamallasalto_id == 10623 or equivalencia.asignaturamallasalto_id == 10627 or equivalencia.asignaturamallasalto_id == 10615:
                                observaciones = recordantiguo.observaciones + " Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15"
                                homologada = True
                            else:
                                observaciones = recordantiguo.observaciones + " Migración con base a RESOLUCIÓN CGA-SO-9-2023-NO15 Y RESOLUCIÓN OCS-SO-18-2023-NO14"
                                homologada = False
                            if not RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                  asignaturamalla=equivalencia.asignaturamallasalto).exists():

                                recordnuevo = RecordAcademico(inscripcion=inscripcion,
                                                              matriculas=recordantiguo.matriculas,
                                                              asignaturamalla=equivalencia.asignaturamallasalto,
                                                              asignatura=equivalencia.asignaturamallasalto.asignatura,
                                                              asignaturaold_id=recordantiguo.asignatura.id,
                                                              nota=recordantiguo.nota,
                                                              asistencia=recordantiguo.asistencia,
                                                              sinasistencia=recordantiguo.sinasistencia,
                                                              fecha=recordantiguo.fecha,
                                                              noaplica=recordantiguo.noaplica,
                                                              aprobada=recordantiguo.aprobada,
                                                              convalidacion=recordantiguo.convalidacion,
                                                              pendiente=recordantiguo.pendiente,
                                                              creditos=equivalencia.asignaturamallasalto.creditos,
                                                              horas=equivalencia.asignaturamallasalto.horas,
                                                              valida=recordantiguo.valida,
                                                              validapromedio=recordantiguo.validapromedio,
                                                              observaciones=observaciones,
                                                              homologada=homologada,
                                                              materiaregular=recordantiguo.materiaregular,
                                                              materiacurso=None,
                                                              completonota=recordantiguo.completonota,
                                                              completoasistencia=recordantiguo.completoasistencia,
                                                              fechainicio=recordantiguo.fechainicio,
                                                              fechafin=recordantiguo.fechafin,
                                                              suficiencia=recordantiguo.suficiencia,
                                                              asignaturamallahistorico_id=recordantiguo.asignaturamalla.id,
                                                              reverso=False)
                                recordnuevo.save()
                                print(u"Crea nuevo record %s" % recordnuevo)


                            elif RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                asignaturamalla=equivalencia.asignaturamallasalto):
                                recordnuevo = RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                             asignaturamalla=equivalencia.asignaturamallasalto)[
                                    0]
                                recordnuevo.matriculas = recordantiguo.matriculas
                                recordnuevo.asignaturamalla = equivalencia.asignaturamallasalto
                                recordnuevo.asignatura = equivalencia.asignaturamallasalto.asignatura
                                recordnuevo.asignaturaold = recordantiguo.asignaturamalla.asignatura
                                recordnuevo.nota = recordantiguo.nota
                                recordnuevo.asistencia = recordantiguo.asistencia
                                recordnuevo.sinasistencia = recordantiguo.sinasistencia
                                recordnuevo.fecha = recordantiguo.fecha
                                recordnuevo.noaplica = recordantiguo.noaplica
                                recordnuevo.aprobada = recordantiguo.aprobada
                                recordnuevo.convalidacion = recordantiguo.convalidacion
                                recordnuevo.pendiente = recordantiguo.pendiente
                                recordnuevo.creditos = equivalencia.asignaturamallasalto.creditos
                                recordnuevo.horas = equivalencia.asignaturamallasalto.horas
                                recordnuevo.valida = recordantiguo.valida
                                recordnuevo.validapromedio = recordantiguo.validapromedio
                                recordnuevo.observaciones = observaciones
                                recordnuevo.homologada = homologada
                                recordnuevo.materiaregular = recordantiguo.materiaregular
                                recordnuevo.materiacurso = None
                                recordnuevo.completonota = recordantiguo.completonota
                                recordnuevo.completoasistencia = recordantiguo.completoasistencia
                                recordnuevo.fechainicio = recordantiguo.fechainicio
                                recordnuevo.fechafin = recordantiguo.fechafin
                                recordnuevo.suficiencia = recordantiguo.suficiencia
                                recordnuevo.asignaturamallahistorico = recordantiguo.asignaturamalla
                                recordnuevo.reverso = False
                                recordnuevo.save()

                            if recordnuevo:
                                historicos = HistoricoRecordAcademico.objects.filter(status=True,
                                                                                     recordacademico=recordantiguo).update(
                                    recordacademico=recordnuevo,
                                    creditos=recordnuevo.creditos,
                                    horas=recordnuevo.horas,
                                    homologada=recordnuevo.homologada)
                                respaldo = RespaldoRecordAcademico.objects.filter(status=True,
                                                                                  recordacademicooriginal=recordantiguo)

                                if not respaldo.exists():
                                    respaldorecord = RespaldoRecordAcademico(
                                        recordacademicooriginal=recordantiguo,
                                        recordacademiconuevo=recordnuevo
                                    )
                                    respaldorecord.save()
                                else:
                                    respaldorecord = respaldo[0]
                                    respaldorecord.recordacademiconuevo = recordnuevo
                                    respaldorecord.save()
                                print(u"Record actualizado %s" % recordnuevo)


                        else:
                            hojadestino.write(fila, 3, "NO ENCONTRO RECORD ANTIGUO %s" % equivalencia.asignaturamalla)
                            fila += 1

                    fila += 1

                    time.sleep(3)

            lin += 1

        libdestino.save(output_folder + libre_origen)
        print(output_folder + libre_origen)
        print("Proceso finalizado. . .")


    except Exception as ex:
        transaction.set_rollback(True)
        print('error: %s' % ex)
        hojadestino.write(fila, 3, str(ex))
        fila += 1
def homologacion_derecho_8():
    # verificar que todos los estudiantes tengan la misma malla
    cadena = ''
    linea, excluidos, conexito = 0, 0, 0
    try:
        libre_origen = '/homologacion_derecho_2_8.xls'
        fuentecabecera = easyxf(
            'font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
        fuentenormal = easyxf(
            'font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
        output_folder = MEDIA_ROOT
        output_folder = os.path.join(os.path.join(BASE_DIR))
        # liborigen = xlrd.open_workbook(output_folder + libre_origen)
        libdestino = xlwt.Workbook()
        hojadestino = libdestino.add_sheet('HOJA1')
        fil = 0
        columnas = [(u"CEDULA", 6000),
                    (u"APELLIDOS Y NOMBRES", 6000),
                    (u"OBSERVACIÓN", 6000)
                    ]
        for col_num in range(len(columnas)):
            hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
            hojadestino.col(col_num).width = columnas[col_num][1]
        fila = 1
        titulacion = 0

        lin = 0
        miarchivo = openpyxl.load_workbook("derecholineaf.xlsx")
        # miarchivo = openpyxl.load_workbook("prueba_2.xlsx")

        ws = miarchivo.get_sheet_by_name("octavo")
        worksheet = ws
        c = 0
        cont = 0
        periodo_id = 177
        carrera_id = 126
        mallaantigua_id = 207
        mallanueva_id = 480

        for row in worksheet.iter_rows(min_row=0):
            if lin >= 0:
                currentValues, cadena = [], ''
                for cell in row:
                    cadena += str(cell.value) + ' '
                    currentValues.append(str(cell.value))
                identificacion = currentValues[0]

                with transaction.atomic():
                    matricula = Matricula.objects.filter(status=True, inscripcion__carrera__id=carrera_id,
                                                         inscripcion__persona__cedula=identificacion).first()
                    cont += 1
                    matricula.pasoayuda = True
                    matricula.save()
                    print(u"%s - %s" % (matricula, cont))
                    inscripcion = matricula.inscripcion
                    hojadestino.write(fila, 0, matricula.inscripcion.persona.identificacion(), fuentenormal)
                    hojadestino.write(fila, 1, matricula.inscripcion.persona.nombre_completo(), fuentenormal)
                    hojadestino.write(fila, 2, matricula.nivel.periodo.nombre, fuentenormal)

                    if InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                       malla_id=mallaantigua_id).exists():
                        imantigua = \
                        InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion, malla_id=mallaantigua_id)[
                            0]
                        imantigua.status = False
                        imantigua.save()
                        print(u"Desactiva antigua inscripcion -----------------------------")

                    if not InscripcionMalla.objects.filter(status=True, inscripcion=inscripcion,
                                                           malla_id=mallanueva_id).exists():
                        imnueva = InscripcionMalla(inscripcion=inscripcion, malla_id=mallanueva_id)
                        imnueva.save()
                        print(u"Crea nueva inscripcion -----------------------------")

                    equivalencias = TablaEquivalenciaAsignaturas.objects.filter(status=True,
                                                                                asignaturamalla__malla_id=mallaantigua_id).order_by(
                        'asignaturamallasalto__nivelmalla__orden')
                    for equivalencia in equivalencias:
                        print(u"nueva - %s" % equivalencia.asignaturamallasalto)
                        recordantiguo = inscripcion.recordacademico_set.filter(status=True,
                                                                               asignaturamalla_id=equivalencia.asignaturamalla_id).first()

                        if recordantiguo:
                            print(u"anterior - %s" % equivalencia.asignaturamalla)
                            print(u"Record antiguo: %s" % recordantiguo)
                            recordnuevo = None
                            recordantiguo.status = False
                            recordantiguo.save(update_asignaturamalla=False)

                            if equivalencia.asignaturamallasalto_id == 10623 or equivalencia.asignaturamallasalto_id == 10627 or equivalencia.asignaturamallasalto_id == 10615:
                                observaciones = recordantiguo.observaciones + " Homologación con base a RESOLUCIÓN CGA-SO-9-2023-NO28 Y RESOLUCIÓN OCS-SO-18-2023-NO15"
                                homologada = True
                            else:
                                observaciones = recordantiguo.observaciones + " Migración con base a RESOLUCIÓN CGA-SO-9-2023-NO15 Y RESOLUCIÓN OCS-SO-18-2023-NO14"
                                homologada = False
                            if not RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                  asignaturamalla=equivalencia.asignaturamallasalto).exists():

                                recordnuevo = RecordAcademico(inscripcion=inscripcion,
                                                              matriculas=recordantiguo.matriculas,
                                                              asignaturamalla=equivalencia.asignaturamallasalto,
                                                              asignatura=equivalencia.asignaturamallasalto.asignatura,
                                                              asignaturaold_id=recordantiguo.asignatura.id,
                                                              nota=recordantiguo.nota,
                                                              asistencia=recordantiguo.asistencia,
                                                              sinasistencia=recordantiguo.sinasistencia,
                                                              fecha=recordantiguo.fecha,
                                                              noaplica=recordantiguo.noaplica,
                                                              aprobada=recordantiguo.aprobada,
                                                              convalidacion=recordantiguo.convalidacion,
                                                              pendiente=recordantiguo.pendiente,
                                                              creditos=equivalencia.asignaturamallasalto.creditos,
                                                              horas=equivalencia.asignaturamallasalto.horas,
                                                              valida=recordantiguo.valida,
                                                              validapromedio=recordantiguo.validapromedio,
                                                              observaciones=observaciones,
                                                              homologada=homologada,
                                                              materiaregular=recordantiguo.materiaregular,
                                                              materiacurso=None,
                                                              completonota=recordantiguo.completonota,
                                                              completoasistencia=recordantiguo.completoasistencia,
                                                              fechainicio=recordantiguo.fechainicio,
                                                              fechafin=recordantiguo.fechafin,
                                                              suficiencia=recordantiguo.suficiencia,
                                                              asignaturamallahistorico_id=recordantiguo.asignaturamalla.id,
                                                              reverso=False)
                                recordnuevo.save()
                                print(u"Crea nuevo record %s" % recordnuevo)


                            elif RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                asignaturamalla=equivalencia.asignaturamallasalto):
                                recordnuevo = RecordAcademico.objects.filter(status=True, inscripcion=inscripcion,
                                                                             asignaturamalla=equivalencia.asignaturamallasalto)[
                                    0]
                                recordnuevo.matriculas = recordantiguo.matriculas
                                recordnuevo.asignaturamalla = equivalencia.asignaturamallasalto
                                recordnuevo.asignatura = equivalencia.asignaturamallasalto.asignatura
                                recordnuevo.asignaturaold = recordantiguo.asignaturamalla.asignatura
                                recordnuevo.nota = recordantiguo.nota
                                recordnuevo.asistencia = recordantiguo.asistencia
                                recordnuevo.sinasistencia = recordantiguo.sinasistencia
                                recordnuevo.fecha = recordantiguo.fecha
                                recordnuevo.noaplica = recordantiguo.noaplica
                                recordnuevo.aprobada = recordantiguo.aprobada
                                recordnuevo.convalidacion = recordantiguo.convalidacion
                                recordnuevo.pendiente = recordantiguo.pendiente
                                recordnuevo.creditos = equivalencia.asignaturamallasalto.creditos
                                recordnuevo.horas = equivalencia.asignaturamallasalto.horas
                                recordnuevo.valida = recordantiguo.valida
                                recordnuevo.validapromedio = recordantiguo.validapromedio
                                recordnuevo.observaciones = observaciones
                                recordnuevo.homologada = homologada
                                recordnuevo.materiaregular = recordantiguo.materiaregular
                                recordnuevo.materiacurso = None
                                recordnuevo.completonota = recordantiguo.completonota
                                recordnuevo.completoasistencia = recordantiguo.completoasistencia
                                recordnuevo.fechainicio = recordantiguo.fechainicio
                                recordnuevo.fechafin = recordantiguo.fechafin
                                recordnuevo.suficiencia = recordantiguo.suficiencia
                                recordnuevo.asignaturamallahistorico = recordantiguo.asignaturamalla
                                recordnuevo.reverso = False
                                recordnuevo.save()

                            if recordnuevo:
                                historicos = HistoricoRecordAcademico.objects.filter(status=True,
                                                                                     recordacademico=recordantiguo).update(
                                    recordacademico=recordnuevo,
                                    creditos=recordnuevo.creditos,
                                    horas=recordnuevo.horas,
                                    homologada=recordnuevo.homologada)
                                respaldo = RespaldoRecordAcademico.objects.filter(status=True,
                                                                                  recordacademicooriginal=recordantiguo)

                                if not respaldo.exists():
                                    respaldorecord = RespaldoRecordAcademico(
                                        recordacademicooriginal=recordantiguo,
                                        recordacademiconuevo=recordnuevo
                                    )
                                    respaldorecord.save()
                                else:
                                    respaldorecord = respaldo[0]
                                    respaldorecord.recordacademiconuevo = recordnuevo
                                    respaldorecord.save()
                                print(u"Record actualizado %s" % recordnuevo)


                        else:
                            hojadestino.write(fila, 3, "NO ENCONTRO RECORD ANTIGUO %s" % equivalencia.asignaturamalla)
                            fila += 1

                    fila += 1

                    time.sleep(3)

            lin += 1

        libdestino.save(output_folder + libre_origen)
        print(output_folder + libre_origen)
        print("Proceso finalizado. . .")


    except Exception as ex:
        transaction.set_rollback(True)
        print('error: %s' % ex)
        hojadestino.write(fila, 3, str(ex))
        fila += 1

print("Función varios")
with concurrent.futures.ThreadPoolExecutor(max_workers=8) as executor:
    from settings import DEBUG
    # if DEBUG:
    #     path_anexo = 'reporte_acortezl__examen_admisión_enero2024.xlsx'
    future_1 = executor.submit(homologacion_derecho_1)
    future_2 = executor.submit(homologacion_derecho_2)
    future_3 = executor.submit(homologacion_derecho_3)
    future_4 = executor.submit(homologacion_derecho_4)
    future_5 = executor.submit(homologacion_derecho_5)
    future_6 = executor.submit(homologacion_derecho_6)
    future_7 = executor.submit(homologacion_derecho_7)
    future_8 = executor.submit(homologacion_derecho_8)