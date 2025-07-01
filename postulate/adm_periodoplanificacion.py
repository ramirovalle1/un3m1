import json
import random
# decoradores
import sys

from functools import reduce
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Count, Max
from django.template.loader import get_template
from django.forms import model_to_dict, TimeInput
from decorators import last_access, secure_module

from django.template import Context
from django.db import transaction
from django.http import HttpResponseRedirect, JsonResponse, HttpResponse
from django.shortcuts import render
from django.db.models.query_utils import Q
from datetime import datetime,date

from openpyxl import workbook as openxl
from openpyxl.styles import Font as openxlFont
from openpyxl.styles.alignment import Alignment as alin

from postulate.models import PeriodoPlanificacion, Convocatoria, Partida, ConvocatoriaTerminosCondiciones, \
    PartidaAsignaturas, ConvocatoriaCalificacion, PartidaTribunal, PersonaAplicarPartida, ConvocatoriaPostulante, \
    PersonaIdiomaPartida, PersonaFormacionAcademicoPartida, PersonaExperienciaPartida, PersonaCapacitacionesPartida, \
    PersonaPublicacionesPartida, CARGOS_TRIBUNAL, RequisitoDocumentoContrato, DetalleCapacitacionPlanificacion, \
    TIEMPO_CAPACITACION, TipoCompetenciaPlanificacion, TipoCompetenciaEspecificaPlanificacion, \
    PeriodoAcademicoConvocatoria, PreguntaPeriodoPlanificacion, DetalleCompetenciaPlanificacion, HistorialPartida, \
    PartidaArmonizacionNomenclaturaTitulo
from postulate.postular import validar_campos
from sagest.models import TipoCompetenciaLaboral, CompetenciaLaboral, DetalleCompetenciaLaboral, NIVEL_COMPETENCIA
from settings import EMAIL_INSTITUCIONAL_AUTOMATICO, ACTUALIZAR_FOTO_ALUMNOS
from sga.commonviews import adduserdata
from postulate.forms import PartidaPlanificacionForm, PeriodoPlanificacionForm, \
    ConvocatoriaForm, PartidaForm, ConvocatoriaTerminosForm, CustomDateInput, TribunalForm, AgendaDisertacionForm, \
    RequisitoDocumentoContratoForm, PartidaPlanificacionAddForm, DetalleCapacitacionPlanificacionForm, \
    TipoCompetenciaPlanificacionForm, \
    TipoCompetenciaEspecificaPlanificacionForm, PeriodoAcademicoConvocatoriaForm, PreguntaPeriodoPlanificacionForm, \
    ValidarPartidaPlanificacionForm
from sga.funciones import log, MiPaginador, numero_a_letras,remover_caracteres_especiales_unicode
from sga.funcionesxhtml2pdf import conviert_html_to_pdf

from sga.models import AreaConocimientoTitulacion, SubAreaConocimientoTitulacion, SubAreaEspecificaConocimientoTitulacion, Carrera, Asignatura, Titulo, Persona
from sga.templatetags.sga_extras import encrypt
from xlwt import *
import xlwt
import xlsxwriter
import io

@login_required(redirect_field_name='ret', login_url='/loginpostulate')
#@secure_module
@last_access
@transaction.atomic()
def view(request):
    global ex
    data = {}
    perfilprincipal = request.session['perfilprincipal']
    persona = request.session['persona']
    periodo = request.session['periodo']
    data['hoy'] = hoy = datetime.now().date()
    data['currenttime'] = datetime.now()
    data['perfil'] = persona.mi_perfil()
    data['periodo'] = periodo

    if request.method == 'POST':
        action = request.POST['action']

        if action == 'addconvocatoria':
            try:
                form = ConvocatoriaForm(request.POST)
                if form.is_valid():
                    convocatoria = Convocatoria(descripcion=form.cleaned_data['descripcion'],
                                                finicio=form.cleaned_data['fechainicio'],
                                                ffin=form.cleaned_data['fechafin'],
                                                tipocontrato=form.cleaned_data['tipocontrato'],
                                                denominacionpuesto=form.cleaned_data['denominacionpuesto'],
                                                modeloevaluativo = form.cleaned_data['modeloevaluativo'],
                                                vigente=form.cleaned_data['vigente'])
                    convocatoria.valtercernivel = form.cleaned_data['valtercernivel']
                    convocatoria.departamento = form.cleaned_data['departamento']
                    convocatoria.valposgrado = form.cleaned_data['valposgrado']
                    convocatoria.valdoctorado = form.cleaned_data['valdoctorado']
                    convocatoria.valcapacitacionmin = form.cleaned_data['valcapacitacionmin']
                    convocatoria.valcapacitacionmax = form.cleaned_data['valcapacitacionmax']
                    convocatoria.valexpdocentemin = form.cleaned_data['valexpdocentemin']
                    convocatoria.valexpdocentemax = form.cleaned_data['valexpdocentemax']
                    convocatoria.valexpadminmin = form.cleaned_data['valexpadminmin']
                    convocatoria.valexpadminmax = form.cleaned_data['valexpadminmax']
                    convocatoria.save(request)
                    log(u'Adicion de convocatoria: %s' % convocatoria, request, "addconvocatoria")
                    return JsonResponse({'result': False, 'mensaje': u'Guardado con exito'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': u'Error al guardar los datos'})

        if action == 'editconvocatoria':
            try:
                form = ConvocatoriaForm(request.POST)
                convocatoria = Convocatoria.objects.get(id=int(encrypt(request.POST['id'])))
                if form.is_valid():
                    convocatoria.descripcion = form.cleaned_data['descripcion']
                    convocatoria.finicio = form.cleaned_data['fechainicio']
                    convocatoria.ffin = form.cleaned_data['fechafin']
                    convocatoria.tipocontrato = form.cleaned_data['tipocontrato']
                    convocatoria.denominacionpuesto = form.cleaned_data['denominacionpuesto']
                    convocatoria.modeloevaluativo = form.cleaned_data['modeloevaluativo']
                    convocatoria.vigente = form.cleaned_data['vigente']
                    convocatoria.valtercernivel = form.cleaned_data['valtercernivel']
                    convocatoria.valposgrado = form.cleaned_data['valposgrado']
                    convocatoria.valdoctorado = form.cleaned_data['valdoctorado']
                    convocatoria.valcapacitacionmin = form.cleaned_data['valcapacitacionmin']
                    convocatoria.valcapacitacionmax = form.cleaned_data['valcapacitacionmax']
                    convocatoria.valexpdocentemin = form.cleaned_data['valexpdocentemin']
                    convocatoria.valexpdocentemax = form.cleaned_data['valexpdocentemax']
                    convocatoria.valexpadminmin = form.cleaned_data['valexpadminmin']
                    convocatoria.valexpadminmax = form.cleaned_data['valexpadminmax']
                    convocatoria.departamento = form.cleaned_data['departamento']
                    convocatoria.save(request)
                    log(u'Edicion de convocatoria: %s' % convocatoria, request, "editconvocatoria")
                    return JsonResponse({'result': False, 'mensaje': 'Edicion Exitosa'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': u'Error al guardar los datos'})

        if action == 'delconvocatoria':
            try:
                with transaction.atomic():
                    convocatoria = Convocatoria.objects.get(id=int(encrypt(request.POST['id'])))
                    convocatoria.status = False
                    convocatoria.save(request)
                    log(u'Elimincion de convocatoria: %s' % convocatoria, request, "delconvocatoria")
                    res_json = {"error": False}
            except Exception as ex:
                res_json = {'error': True, "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)

        if action == 'deltribunalsegundaetapa':
            try:
                with transaction.atomic():
                    convocatoria = Convocatoria.objects.get(id=int(request.POST['id']))
                    for trib in PartidaTribunal.objects.filter(status=True, partida__convocatoria=convocatoria, tipo=2):
                        trib.status = False
                        trib.save(request)
                    convocatoria.save(request)
                    log(u'Elimincion de tribunal segunda etapa: %s' % convocatoria, request, "deltribunalsegundaetapa")
                    res_json = {"error": False}
            except Exception as ex:
                res_json = {'error': True, "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)

        if action == 'deltribunalprimeraetapa':
            try:
                with transaction.atomic():
                    convocatoria = Convocatoria.objects.get(id=int(request.POST['id']))
                    for trib in PartidaTribunal.objects.filter(status=True, partida__convocatoria=convocatoria, tipo=1):
                        trib.status = False
                        trib.save(request)
                    convocatoria.save(request)
                    log(u'Elimincion de tribunal segunda etapa: %s' % convocatoria, request, "deltribunalsegundaetapa")
                    res_json = {"error": False}
            except Exception as ex:
                res_json = {'error': True, "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)

        if action == 'editvigente':
            try:
                convocatoria = Convocatoria.objects.get(id=int(encrypt(request.POST['id'])))
                convocatoria.vigente = request.POST['vigente']
                convocatoria.save(request)
                log(u'Edicion de estado Vigencia de convocatoria: %s' % convocatoria, request, "edit")
                return JsonResponse({'result': 'ok', 'mensaje': 'Edicion Exitosa'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', 'mensaje': u'Error al guardar los datos'})

        if action == 'editapelacion':
            try:
                convocatoria = Convocatoria.objects.get(id=int(encrypt(request.POST['id'])))
                convocatoria.apelacion = request.POST['apelacion']
                convocatoria.save(request)
                log(u'Edicion de estado Apelación de convocatoria: %s' % convocatoria, request, "edit")
                return JsonResponse({'result': 'ok', 'mensaje': 'Edicion Exitosa'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', 'mensaje': u'Error al guardar los datos'})

        if action == 'editsegundaetapa':
            try:
                convocatoria = Convocatoria.objects.get(id=int(encrypt(request.POST['id'])))
                convocatoria.segundaetapa = request.POST['segundaetapa']
                convocatoria.save(request)
                log(u'Edicion de estado Disertación/Apelación de convocatoria: %s' % convocatoria, request, "edit")
                return JsonResponse({'result': 'ok', 'mensaje': 'Edicion Exitosa'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', 'mensaje': u'Error al guardar los datos'})

        if action == 'addperiodoplanificacion':
            try:
                data['title'] = u'Adicionar Periodo de planificación'
                form = PeriodoPlanificacionForm(request.POST)
                if form.is_valid():
                    periodoplanificacion = PeriodoPlanificacion(nombre=form.cleaned_data['nombre'],
                                                                finicio=form.cleaned_data['fechainicio'],
                                                                ffin=form.cleaned_data['fechafin'],
                                                                vigente=form.cleaned_data['vigente'])
                    periodoplanificacion.save(request)
                    log(u'Adicion de periodo de planificacion: %s' % periodoplanificacion, request, "addperiodoplanificacion")
                    return JsonResponse({'result': False, 'mensaje': u'Guardado con exito'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': u'Error al guardar los datos'})

        if action == 'editperiodoplanificacion':
            try:
                form = PeriodoPlanificacionForm(request.POST)
                pplanificacion = PeriodoPlanificacion.objects.get(id=int(encrypt(request.POST['id'])))
                if form.is_valid():
                    pplanificacion.nombre = form.cleaned_data['nombre']
                    pplanificacion.finicio = form.cleaned_data['fechainicio']
                    pplanificacion.ffin = form.cleaned_data['fechafin']
                    pplanificacion.vigente = form.cleaned_data['vigente']
                    pplanificacion.save(request)
                    log(u'Edicion de periodo de planificacion: %s' % pplanificacion, request, "editperiodoplanificacion")
                    return JsonResponse({'result': False, 'mensaje': 'Edicion Exitosa'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': u'Error al guardar los datos'})

        if action == 'delperiodoplanificacion':
            try:
                with transaction.atomic():
                    pplanificacion = PeriodoPlanificacion.objects.get(id=int(encrypt(request.POST['id'])))
                    pplanificacion.status = False
                    pplanificacion.save(request)
                    log(u'Elimincion de periodo de planificacion: %s' % pplanificacion, request, "delperiodoplanificacion")
                    res_json = {"error": False}
            except Exception as ex:
                res_json = {'error': True, "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)

        if action == 'editvigentepplanificacion':
            try:
                pplanificacion = PeriodoPlanificacion.objects.get(id=int(encrypt(request.POST['id'])))
                pplanificacion.vigente = request.POST['vigente']
                pplanificacion.save(request)
                log(u'Edicion de vigencia de periodo de planificacion: %s' % pplanificacion, request, "editvigentepplanificacion")
                return JsonResponse({'result': 'ok', 'mensaje': 'Edicion Exitosa'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', 'mensaje': u'Error al guardar los datos'})

        if action == 'addpartida':
            try:
                form = PartidaForm(request.POST)
                idc = int(encrypt(request.POST['idc']))
                convocatoria = Convocatoria.objects.get(id=idc)
                if form.is_valid():
                    partida = Partida(
                        convocatoria_id=idc,
                        codpartida=form.cleaned_data['codpartida'],
                        titulo=form.cleaned_data['titulo'],
                        descripcion=form.cleaned_data['descripcion'],
                        carrera=form.cleaned_data['carrera'],
                        nivel=form.cleaned_data['nivel'],
                        modalidad=form.cleaned_data['modalidad'],
                        dedicacion=form.cleaned_data['dedicacion'],
                        jornada=form.cleaned_data['jornada'],
                        rmu=form.cleaned_data['rmu'],
                        vigente=form.cleaned_data['vigente'])
                    partida.save()
                    for ca in form.cleaned_data['campoamplio']:
                        partida.campoamplio.add(ca)
                    for caesp in form.cleaned_data['campoespecifico']:
                        partida.campoespecifico.add(caesp)
                    for campdet in form.cleaned_data['campodetallado']:
                        partida.campodetallado.add(campdet)
                    for tit in form.cleaned_data['titulos']:
                        partida.titulos.add(tit)
                    partida.save(request)

                    for asignatura in form.cleaned_data['asignatura']:
                        partidaasignatura = PartidaAsignaturas(partida=partida,
                                                               asignatura=asignatura)
                        partidaasignatura.save(request)
                    log(u'Adicion de partida y partida asignatura: %s' % partida, request, "addpartida")
                    return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', 'mensaje': u'Error al guardar los datos'})

        if action == 'editpartida':
            try:
                form = PartidaForm(request.POST)
                partida = Partida.objects.get(id=request.POST['id'])
                partidaasignaturas = PartidaAsignaturas.objects.filter(partida=partida, status=True)
                if form.is_valid():
                    partida.codpartida = form.cleaned_data['codpartida']
                    # partida.titulo = form.cleaned_data['titulo']
                    # partida.descripcion = form.cleaned_data['descripcion']
                    partida.carrera = form.cleaned_data['carrera']
                    partida.denominacionpuesto = form.cleaned_data['denominacionpuesto']
                    partida.nivel = form.cleaned_data['nivel']
                    partida.modalidad = form.cleaned_data['modalidad']
                    partida.dedicacion = form.cleaned_data['dedicacion']
                    partida.jornada = form.cleaned_data['jornada']
                    partida.rmu = form.cleaned_data['rmu']
                    partida.vigente = form.cleaned_data['vigente']
                    partida.save(request)

                    partida.campoamplio.clear()
                    partida.campoespecifico.clear()
                    partida.campodetallado.clear()
                    partida.titulos.clear()
                    for ca in form.cleaned_data['campoamplio']:
                        partida.campoamplio.add(ca)
                    for caesp in form.cleaned_data['campoespecifico']:
                        partida.campoespecifico.add(caesp)
                    for campdet in form.cleaned_data['campodetallado']:
                        partida.campodetallado.add(campdet)
                    for tit in form.cleaned_data['titulos']:
                        partida.titulos.add(tit)
                    partida.save(request)

                    for partidaasignatura in partidaasignaturas:
                        partidaasignatura.status = False
                        partidaasignatura.save(request)

                    for asignatura in form.cleaned_data['asignatura']:
                        if not PartidaAsignaturas.objects.filter(partida=partida, asignatura=asignatura, status=True).exists():
                            partidasave = PartidaAsignaturas(partida=partida,
                                                             asignatura=asignatura, )
                            partidasave.save(request)
                    log(u'Edicion de partida y partidaasignatura: %s' % partida, request, "editpartida")
                    return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', 'mensaje': u'Error al guardar los datos'})

        if action == 'delpartida':
            try:
                with transaction.atomic():
                    partida = Partida.objects.get(id=int(encrypt(request.POST['id'])))
                    partida.status = False
                    partida.save(request)

                    for partidaasignatura in PartidaAsignaturas.objects.filter(partida=partida, status=True):
                        partidaasignatura.status = False
                        partidaasignatura.save(request)

                    log(u'Elimincion de partida: %s' % partida, request, "delpartida")
                    res_json = {"error": False}
            except Exception as ex:
                res_json = {'error': True, "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)

        if action == 'editvigentepartida':
            try:
                partida = Partida.objects.get(id=int(encrypt(request.POST['id'])))
                partida.vigente = request.POST['vigente']
                partida.save(request)
                log(u'Edicion de vigencia de partida: %s' % partida, request, "editvigentepartida")
                return JsonResponse({'result': 'ok', 'mensaje': 'Edicion Exitosa'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', 'mensaje': u'Error al guardar los datos'})

        if action == 'editvigentefirma':
            try:
                partida = PartidaTribunal.objects.get(id=int(encrypt(request.POST['id'])))
                partida.firma = request.POST['firma']
                partida.save(request)
                log(u'Edicion de firma de tritunal: %s' % partida, request, "editvigentefirma")
                return JsonResponse({'result': 'ok', 'mensaje': 'Edicion Exitosa'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', 'mensaje': u'Error al guardar los datos'})

        if action == 'addpartidaplanificacion':
            try:
                form = PartidaPlanificacionAddForm(request.POST)
                idc = int(encrypt(request.POST['idc']))
                action = request.POST['action']
                periodoplanificacion = PeriodoPlanificacion.objects.get(id=idc)
                if form.is_valid():
                    partida = Partida(
                        periodoplanificacion_id=idc,
                        codpartida=form.cleaned_data['codpartida'],
                        # titulo=form.cleaned_data['titulo'],
                        denominacionpuesto=form.cleaned_data['denominacionpuesto'],
                        # descripcion=form.cleaned_data['descripcion'],
                        carrera=form.cleaned_data['carrera'],
                        nivel=form.cleaned_data['nivel'],
                        modalidad=form.cleaned_data['modalidad'],
                        dedicacion=form.cleaned_data['dedicacion'],
                        jornada=form.cleaned_data['jornada'],
                        rmu=form.cleaned_data['rmu'],
                        estado=form.cleaned_data['estado'],
                        temadisertacion=form.cleaned_data['temadisertacion'],
                        observacion=form.cleaned_data['observacion'])
                    partida.save()
                    for ca in form.cleaned_data['campoamplio']:
                        partida.campoamplio.add(ca)
                    for caesp in form.cleaned_data['campoespecifico']:
                        partida.campoespecifico.add(caesp)
                    for campdet in form.cleaned_data['campodetallado']:
                        partida.campodetallado.add(campdet)
                    for tit in form.cleaned_data['titulos']:
                        partida.titulos.add(tit)
                    partida.save(request)

                    for asignatura in form.cleaned_data['asignatura']:
                        partidaasignatura = PartidaAsignaturas(partida=partida,
                                                               asignatura=asignatura)
                        partidaasignatura.save(request)
                    datosdetalle = request.POST.getlist('infoDetalle[]')
                    datoscompetencia = request.POST.getlist('infoDetalle2[]')

                    datosdetalle = [datosdetalle[li:li + 5] for li in range(0, datosdetalle.__len__(), 5)]

                    for det in datosdetalle:
                        if not det[0] or not det[1] or not det[2] or not det[3] or not det[4]:
                            raise NameError(f'Complete todos los campos de requeridos en criterios seleccionados.')
                        detalle = DetalleCapacitacionPlanificacion(
                            partida=partida,
                            tipocompetencia_id=int(det[0]),
                            tiempocapacitacion=int(det[1]),
                            canttiempocapacitacion=int(det[2]),
                            cespecifica_id=int(det[3]),
                            descripcioncapacitacion=det[4],
                            )
                        detalle.save(request)

                    for det in datoscompetencia:
                        competencia = DetalleCompetenciaPlanificacion(
                            partida=partida,
                            competencialaboral_id=int(det))
                        competencia.save(request)
                    for arm in form.cleaned_data['armonizacion']:
                        partarmonizacion = PartidaArmonizacionNomenclaturaTitulo(
                                combinacion=arm,
                                partida=partida
                            )
                        partarmonizacion.save(request)
                log(u'Adicion de partida de planificacion y partida asignatura y detalle de partida: %s' % partida, request, "addpartidaplanificacion")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', 'mensaje': f'Error: {ex}'})

        if action == 'editpartidaplanificacion':
            try:
                form = PartidaPlanificacionAddForm(request.POST)
                partida = Partida.objects.get(id=request.POST['id'])
                detallecapacitaciones = DetalleCapacitacionPlanificacion.objects.filter(partida=partida, status=True)
                partidaasignaturas = PartidaAsignaturas.objects.filter(partida=partida, status=True)
                exclude_borrar_armonizacion = []
                if form.is_valid():
                    partida.codpartida = form.cleaned_data['codpartida']
                    # partida.titulo = form.cleaned_data['titulo']
                    partida.denominacionpuesto = form.cleaned_data['denominacionpuesto']
                    # partida.descripcion = form.cleaned_data['descripcion']
                    partida.carrera = form.cleaned_data['carrera']
                    partida.nivel = form.cleaned_data['nivel']
                    partida.modalidad = form.cleaned_data['modalidad']
                    partida.dedicacion = form.cleaned_data['dedicacion']
                    partida.jornada = form.cleaned_data['jornada']
                    partida.rmu = form.cleaned_data['rmu']
                    partida.estado = form.cleaned_data['estado']
                    partida.temadisertacion = form.cleaned_data['temadisertacion']
                    partida.observacion = form.cleaned_data['observacion']
                    partida.save(request)

                    partida.campoamplio.clear()
                    partida.campoespecifico.clear()
                    partida.campodetallado.clear()
                    partida.titulos.clear()
                    for ca in form.cleaned_data['campoamplio']:
                        partida.campoamplio.add(ca)
                    for caesp in form.cleaned_data['campoespecifico']:
                        partida.campoespecifico.add(caesp)
                    for campdet in form.cleaned_data['campodetallado']:
                        partida.campodetallado.add(campdet)
                    for tit in form.cleaned_data['titulos']:
                        partida.titulos.add(tit)
                    partida.save(request)

                    for partidaasignatura in partidaasignaturas:
                        partidaasignatura.status = False
                        partidaasignatura.save(request)

                    for detallecapacitacion in detallecapacitaciones:
                         detallecapacitacion.status = False
                         detallecapacitacion.save(request)

                    for asignatura in form.cleaned_data['asignatura']:
                        if not PartidaAsignaturas.objects.filter(partida=partida, asignatura=asignatura, status=True).exists():
                            partidasave = PartidaAsignaturas(partida=partida,
                                                             asignatura=asignatura, )
                            partidasave.save(request)
                    datosdetalle = request.POST.getlist('infoDetalle[]')
                    datoscompetencia = request.POST.getlist('infoDetalle2[]')

                    datosdetalle = [datosdetalle[li:li + 5] for li in range(0, datosdetalle.__len__(), 5)]
                    listaidcapacitacion = []

                    for det in datosdetalle:
                        detalle = DetalleCapacitacionPlanificacion.objects.filter(status=True, partida=partida,
                                                                        tipocompetencia_id=int(det[0]),
                                                                        tiempocapacitacion=int(det[1]),
                                                                        cespecifica_id=int(det[3]))
                        if detalle:
                            detalle = detalle[0]
                            detalle.tipocompetencia_id=int(det[0])
                            detalle.tiempocapacitacion=int(det[1])
                            detalle.canttiempocapacitacion=int(det[2])
                            detalle.cespecifica_id=int(det[3])
                            detalle.descripcioncapacitacion=(det[4])

                        else:
                            detalle = DetalleCapacitacionPlanificacion(
                                partida=partida,
                                tipocompetencia_id=int(det[0]),
                                tiempocapacitacion=int(det[1]),
                                canttiempocapacitacion=int(det[2]),
                                cespecifica_id=int(det[3]),
                                descripcioncapacitacion=(det[4]),
                            )
                        detalle.save(request)
                        listaidcapacitacion.append(detalle.id)

                    eliminar = DetalleCapacitacionPlanificacion.objects.filter(status=True, partida=partida).exclude(id__in=listaidcapacitacion)
                    for detalle in eliminar:
                        detalle.status = False
                        detalle.save(request)
                        log(u'Elimina detalle de partida: %s' % detalle, request, "del")

                    listaidcompetencia = []

                    for det in datoscompetencia:
                        competencia = DetalleCompetenciaPlanificacion.objects.filter(partida=partida,
                                                                        competencialaboral_id=int(det))
                        if competencia:
                            competencia = competencia[0]
                            competencia.competencialaboral_id=int(det)
                        else:
                            competencia = DetalleCompetenciaPlanificacion(
                                partida=partida,
                                competencialaboral_id=int(det))
                        competencia.save(request)
                        listaidcompetencia.append(competencia.id)

                    eliminar = DetalleCompetenciaPlanificacion.objects.filter(status=True, partida=partida).exclude(id__in=listaidcompetencia)
                    for competencia in eliminar:
                        competencia.status = False
                        competencia.save(request)
                        log(u'Elimina competencia de partida: %s' % competencia, request, "del")

                    # if datosdetalle:
                    #     c = 0
                    #     while c < len(datosdetalle):
                    #         detalle = DetalleCapacitacionPlanificacion(
                    #             partida=partida,
                    #             tipocompetencia_id=int(datosdetalle[c]),
                    #             tiempocapacitacion=int(datosdetalle[c + 1]),
                    #             canttiempocapacitacion=int(datosdetalle[c + 2]),
                    #             cespecifica_id=int(datosdetalle[c + 3]) if int(datosdetalle[c + 3])!= 0 else None,
                    #             descripcioncapacitacion=datosdetalle[c + 4],
                    #         )
                    #         detalle.save(request)
                    #         c += 5
                    #
                    # else:
                    #     eliminar = DetalleCapacitacionPlanificacion.objects.filter(status=True, partida=partida)
                    #     for detalle in eliminar:
                    #         detalle.status = False
                    #         detalle.save(request)
                    #         log(u'Elimina detalle de partida: %s' % detalle, request, "del")
                    historial = HistorialPartida(partida=partida,
                                          estado=partida.estado,
                                          persona=persona,
                                          observacion='Editó partida en %s' %(str(datetime.now().date())))
                    historial.save()

                    for arm in form.cleaned_data['armonizacion']:
                        if not PartidaArmonizacionNomenclaturaTitulo.objects.filter(status=True, combinacion=arm,partida=partida).exists():
                            partarmonizacion = PartidaArmonizacionNomenclaturaTitulo(
                                combinacion=arm,
                                partida=partida
                            )
                            partarmonizacion.save(request)
                            exclude_borrar_armonizacion.append(partarmonizacion.id)
                        else:
                            partarmonizacion = PartidaArmonizacionNomenclaturaTitulo.objects.get(status=True, combinacion=arm,partida=partida)
                            exclude_borrar_armonizacion.append(partarmonizacion.id)
                    armo_delete = PartidaArmonizacionNomenclaturaTitulo.objects.filter(status=True,partida=partida).exclude(id__in=exclude_borrar_armonizacion)
                    for dele in armo_delete:
                        dele.status = False
                        dele.save(request)
                    for partidaasignatura in partidaasignaturas:
                        partidaasignatura.status = False
                        partidaasignatura.save(request)

                    log(u'Edicion de partida y partidaasignatura: %s' % partida, request, "editpartidaplanificacion")
                    return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', 'mensaje': u'Error al guardar los datos'})

        if action == 'bloquear_subtipo':
            try:
                filtro = TipoCompetenciaPlanificacion.objects.get(id=request.POST['id'])
                if filtro.aplicasubtipo:
                    return JsonResponse({"result": 'ok'})
                else:
                    return JsonResponse({"result": 'bad'})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})


        if action == 'traercompetencia':
            try:
                filtro = TipoCompetenciaLaboral.objects.get(id=request.POST['id'])
                competenciasform= CompetenciaLaboral.objects.filter(tipo=filtro, status=True)
                resp = [{'id': qs.pk, 'text': f"{qs.denominacion}" } for qs in competenciasform]
                return JsonResponse({'result': 'ok', 'results':(resp)})

            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        if action == 'traernivel':
            try:
                filtro = CompetenciaLaboral.objects.get(id=request.POST['id'])
                nivelform= DetalleCompetenciaLaboral.objects.filter(competencia=filtro, status=True)
                resp = [{'id': qs.pk, 'text': f"{qs.get_nivel_display()}" } for qs in nivelform]
                return JsonResponse({'result': 'ok', 'results':(resp)})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        if action == 'delpartidaplanificacion':
            try:
                with transaction.atomic():
                    partida = Partida.objects.get(id=int(encrypt(request.POST['id'])))
                    partida.status = False
                    partida.save(request)
                    for partidaasignatura in PartidaAsignaturas.objects.filter(partida=partida, status=True):
                        partidaasignatura.status = False
                        partidaasignatura.save(request)
                    for detalleplanificacion in DetalleCapacitacionPlanificacion.objects.filter(partida=partida, status=True):
                        detalleplanificacion.status = False
                        detalleplanificacion.save(request)

                    log(u'Eliminacion de partida de planificacion: %s' % partida, request, "delpartidaplanificacion")
                    res_json = {"error": False}
            except Exception as ex:
                res_json = {'error': True, "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)

        if action == 'deldetallecapacitacion':
            try:
                with transaction.atomic():
                    detallecapacitacion = DetalleCapacitacionPlanificacion.objects.get(id=int(encrypt(request.POST['iddetalle'])))
                    detallecapacitacion.status = False
                    detallecapacitacion.save(request)

                    log(u'Eliminacion de detalle de partida de planificacion: %s' % detallecapacitacion, request, "deldetallecapacitacion")
                    res_json = {"error": False}
            except Exception as ex:
                res_json = {'error': True, "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)

        # if action == 'revisarpartidaplanificacion':
        #     with transaction.atomic():
        #         try:
        #             partida = Partida.objects.get(pk=int(request.POST['id']))
        #             historialbase = HistorialAprobacionPartida.objects.filter(partida=partida).last()
        #             f = HistorialAprobacionPartidaForm(request.POST)
        #             if f.is_valid():
        #                 partida.estado = int(f.cleaned_data['estado'])
        #                 partida.save(request)
        #                 log(u'Se revisó la partida de planificacion: %s' % partida, request, "edit")
        #                 nota = HistorialAprobacionPartida(partida=partida,
        #                                           observacion=f.cleaned_data['observacion'],
        #                                           estado=int(f.cleaned_data['estado']))
        #                 nota.save(request)
        #                 log(u'Se guard{o historial de aprobación la partida de planificacion: %s' % nota, request, "add")
        #                 return JsonResponse({"result": False}, safe=False)
        #             else:
        #                 transaction.set_rollback(True)
        #                 return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)
        #         except Exception as ex:
        #             transaction.set_rollback(True)
        #             return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        if action == 'editestadopartida':
            try:
                partida = Partida.objects.get(id=int(encrypt(request.POST['id'])))
                partida.estado = request.POST['estado']
                partida.save(request)
                log(u'Edicion de estado de partida de planificacion: %s' % partida, request, "editestadopartida")
                return JsonResponse({'result': 'ok', 'mensaje': 'Edicion Exitosa'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', 'mensaje': u'Error al guardar los datos'})

        if action == 'addtermino':
            try:
                form = ConvocatoriaTerminosForm(request.POST)
                if form.is_valid():
                    termino = ConvocatoriaTerminosCondiciones(descripcion=form.cleaned_data['descripcion'], convocatoria_id=int(encrypt(request.POST['id'])))
                    termino.save(request)
                    log(u'Adicion Terminos y Condiciones: %s' % termino, request, "addtermino")
                    return JsonResponse({'result': False, 'mensaje': u'Guardado con exito'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': u'Error al guardar los datos'})

        if action == 'edittermino':
            try:
                form = ConvocatoriaTerminosForm(request.POST)
                termino = ConvocatoriaTerminosCondiciones.objects.get(id=int(encrypt(request.POST['id'])))
                if form.is_valid():
                    termino.descripcion = form.cleaned_data['descripcion']
                    termino.save(request)
                    log(u'Edicion de Termino y Condiciones: %s' % termino, request, "edittermino")
                    return JsonResponse({'result': False, 'mensaje': 'Edicion Exitosa'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': u'Error al guardar los datos'})

        if action == 'deltermino':
            try:
                with transaction.atomic():
                    termino = ConvocatoriaTerminosCondiciones.objects.get(id=int(encrypt(request.POST['id'])))
                    termino.status = False
                    termino.save(request)
                    log(u'Elimincion de Terminos y Condiciones: %s' % termino, request, "deltermino")
                    res_json = {"error": False}
            except Exception as ex:
                res_json = {'error': True, "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)

        # ADICIONAR PERSONAS PARA CONFIGURACION DE TRIBUNALES

        if action == 'addtribunal':
            try:
                dat = request.POST
                partidas = request.POST.getlist('partida')
                if len(partidas) > 1:
                    for p in partidas:
                        part = Partida.objects.get(id=p)
                        if PartidaTribunal.objects.filter(status=True, partida=part, cargos=int(dat['cargo']), tipo=int(dat['tipo'])).exists():
                            return JsonResponse({'result': True, "mensaje": u"Cargo ya ocupado en la partida " + str(part.__str__())})
                        if PartidaTribunal.objects.filter(status=True, partida=part, persona_id=int(dat['persona']), tipo=int(dat['tipo'])).exists():
                            return JsonResponse({'result': True, "mensaje": u"Esta persona ya existe en la partida " + str(part.__str__())})
                        ptribunal = PartidaTribunal(partida_id=int(p), persona_id=int(dat['persona']), cargos=int(dat['cargo']), tipo=int(dat['tipo']))
                        if 'firma' in dat:
                            ptribunal.firma = True if dat['firma'] == 'on' else False
                        ptribunal.save(request)
                        log(u'Agrego personas al tribunal: %s' % ptribunal, request, "add")
                else:
                    part = Partida.objects.get(id=partidas[0])
                    if PartidaTribunal.objects.filter(status=True, partida=part, cargos=int(dat['cargo']), tipo=int(dat['tipo'])).exists():
                        return JsonResponse({'result': True, "mensaje": u"Cargo ya ocupado en la partida " + str(part.__str__())})
                    if PartidaTribunal.objects.filter(status=True, partida=part, persona_id=int(dat['persona']), tipo=int(dat['tipo'])).exists():
                        return JsonResponse({'result': True, "mensaje": u"Esta persona ya existe en la partida " + str(part.__str__())})
                    ptribunal = PartidaTribunal(partida_id=int(partidas[0]), persona_id=int(dat['persona']), cargos=int(dat['cargo']), tipo=int(dat['tipo']))
                    if 'firma' in dat:
                        ptribunal.firma = True if dat['firma'] == 'on' else False
                    ptribunal.save(request)
                    log(u'Agrego personas al tribunal: %s' % ptribunal, request, "add")
                return JsonResponse({'result': False, 'mensaje': 'Edicion Exitosa'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': str(ex)})

        if action == 'delmiembrotribunal':
            try:
                with transaction.atomic():
                    partida = PartidaTribunal.objects.get(id=int(encrypt(request.POST['id'])))
                    partida.status = False
                    partida.save(request)
                    log(u'Elimincion miembro de tribunal: %s' % partida, request, "del")
                    res_json = {"error": False}
            except Exception as ex:
                res_json = {'error': True, "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)

        if action == 'agendardisertacion':
            try:
                form = AgendaDisertacionForm(request.POST)
                if form.is_valid():
                    postulacion = PersonaAplicarPartida.objects.get(pk=request.POST['idp'])
                    if 'id' in request.POST:
                        convocatoria = ConvocatoriaPostulante.objects.get(id=int(encrypt(request.POST['id'])))
                    else:
                        convocatoria = ConvocatoriaPostulante(persona=postulacion)
                    convocatoria.tema = form.cleaned_data['tema']
                    convocatoria.fechaasistencia = form.cleaned_data['fechaasistencia']
                    convocatoria.horasistencia = form.cleaned_data['horasistencia']
                    convocatoria.lugar = form.cleaned_data['lugar']
                    convocatoria.observacion = form.cleaned_data['observacion']
                    convocatoria.save(request)
                    log(u'Edicion de Agenda Convocatoria Disertación: %s' % convocatoria, request, "add")
                    return JsonResponse({'result': False, 'mensaje': 'Edicion Exitosa'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': u'Error al guardar los datos'})

        elif action == 'addrequisito':
            try:
                form = RequisitoDocumentoContratoForm(request.POST,request.FILES)
                if not form.is_valid():
                    raise NameError(str([{k: v[0]} for k, v in form.errors.items()][0]))
                archivo = None
                if 'archivo' in request.FILES:
                    archivo = request.FILES['archivo']
                    archivo._name = remover_caracteres_especiales_unicode(archivo._name)
                    ext = archivo._name.split('.')[-1]
                    if not ext in ['pdf']:
                        raise NameError(u'Formato de archivo incorrecto, subir en formato .pdf')
                    if archivo.size > 10194304:
                        raise NameError(u'Tamaño maximo permitod es 10Mb')
                anio = form.cleaned_data['anio']
                requisito = RequisitoDocumentoContrato(
                    nombre = form.cleaned_data['nombre'],
                    descripcion = form.cleaned_data['descripcion'],
                    anio = anio.date(),
                    obligatorio = form.cleaned_data['obligatorio'],
                    activo=form.cleaned_data['activo'],
                    tipo = form.cleaned_data['tipo'],
                    archivo=archivo
                )
                requisito.save(request)
                log(u'Agregó requisito: %s'%(requisito),request,'add')
                return JsonResponse({'result':False, 'mensaje': 'Registro guardado'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': u'Error al guardar los datos. Detalle: %s'%(ex.__str__())})

        elif action == 'addticompeplanificacion':
            try:
                nombre = request.POST['nombre']
                aplica = True if 'aplicasubtipo' in request.POST else False
                filtro = TipoCompetenciaPlanificacion(nombre=nombre, aplicasubtipo=aplica)
                filtro.save(request)
                log(u'Registro tipo de competencia de planificación: %s' % filtro, request, "add")
                return JsonResponse({'result': False, 'mensaje': 'Registro Exitoso'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': f'Error al guardar los datos. {ex}'})

        elif action == 'editticompeplanificacion':
            try:
                form = TipoCompetenciaPlanificacionForm(request.POST)
                filtro = TipoCompetenciaPlanificacion.objects.get(id=int(encrypt(request.POST['id'])))
                if form.is_valid():
                    filtro.nombre = form.cleaned_data['nombre']
                    filtro.aplicasubtipo = form.cleaned_data['aplicasubtipo']
                    filtro.save(request)
                    log(u'Edito tipo de competencia de planificación: %s' % filtro, request, action)
                    return JsonResponse({'result': False, 'mensaje': 'Edicion Exitosa'})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': f'Error al guardar los datos. {ex}'})

        elif action == 'addticompeespeplanificacion':
            try:
                nombre = request.POST['nombre']
                filtro = TipoCompetenciaEspecificaPlanificacion(nombre=nombre)
                filtro.save(request)
                log(u'Registro tipo de competencia específica de planificación: %s' % filtro, request, "add")
                return JsonResponse({'result': False, 'mensaje': 'Registro Exitoso'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': f'Error al guardar los datos. {ex}'})

        elif action == 'editticompeespeplanificacion':
            try:
                form = TipoCompetenciaEspecificaPlanificacionForm(request.POST)
                filtro = TipoCompetenciaEspecificaPlanificacion.objects.get(id=int(encrypt(request.POST['id'])))
                if form.is_valid():
                    filtro.nombre = form.cleaned_data['nombre']
                    filtro.save(request)
                    log(u'Edito tipo de competencia específica de planificación: %s' % filtro, request, action)
                    return JsonResponse({'result': False, 'mensaje': 'Edicion Exitosa'})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': f'Error al guardar los datos. {ex}'})

        elif action == 'editobligatorio':
            try:
                if not 'id' in request.POST:
                    raise NameError()
                id = int(encrypt(request.POST['id']))
                requi = RequisitoDocumentoContrato.objects.get(pk=id)
                requi.obligatorio = request.POST['obligatorio']
                requi.save(request)
                log('Edito el esto de obligatoriedad %s'%(requi.__str__()),request,'edit')
                return JsonResponse({'result':False, 'mensaje': 'Registro editado'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': u'Error al guardar los datos. Detalle: %s'%(ex.__str__())})

        elif action == 'editactivo':
            try:
                if not 'id' in request.POST:
                    raise NameError()
                id = int(encrypt(request.POST['id']))
                requi = RequisitoDocumentoContrato.objects.get(pk=id)
                requi.activo = request.POST['activo']
                requi.save(request)
                log('Edito el esto de obligatoriedad %s'%(requi.__str__()),request,'edit')
                return JsonResponse({'result':False, 'mensaje': 'Registro editado'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': u'Error al guardar los datos. Detalle: %s'%(ex.__str__())})

        elif action == 'editrequisito':
            try:
                form = RequisitoDocumentoContratoForm(request.POST, request.FILES)
                requisito = RequisitoDocumentoContrato.objects.get(id =int(encrypt(request.POST['id'])))
                if not form.is_valid():
                    raise NameError(str([{k: v[0]} for k, v in form.errors.items()][0]))
                archivo = None
                if 'archivo' in request.FILES:
                    archivo = request.FILES['archivo']
                    archivo._name = remover_caracteres_especiales_unicode(archivo._name)
                    ext = archivo._name.split('.')[-1]
                    if not ext in ['pdf']:
                        raise NameError(u'Formato de archivo incorrecto, subir en formato .pdf')
                    if archivo.size > 10194304:
                        raise NameError(u'Tamaño maximo permitod es 10Mb')

                    requisito.archivo = archivo
                anio = form.cleaned_data['anio']
                requisito.nombre = form.cleaned_data['nombre']
                requisito.descripcion = form.cleaned_data['descripcion']
                requisito.anio = anio.date()
                requisito.obligatorio = form.cleaned_data['obligatorio']
                requisito.activo = form.cleaned_data['activo']
                requisito.tipo = form.cleaned_data['tipo']
                requisito.save(request)
                log(u'Editó requisito: %s' % (requisito), request, 'edit')
                return JsonResponse({'result': False, 'mensaje': 'Registro guardado'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': f'Error al guardar los datos. Detalle: {ex.__str__()}'})

        if action == 'deleteticompeplanificacion':
            try:
                filtro = TipoCompetenciaPlanificacion.objects.get(id=int(encrypt(request.POST['id'])))
                filtro.status = False
                filtro.save(request)
                log(u'Elimino tipo de competencia de planificación: %s' % filtro, request, "edit")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        if action == 'deleteticompeespeplanificacion':
            try:
                filtro = TipoCompetenciaEspecificaPlanificacion.objects.get(id=int(encrypt(request.POST['id'])))
                filtro.status = False
                filtro.save(request)
                log(u'Elimino tipo de competencia específica de planificación: %s' % filtro, request, "delete")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        if action == 'addpregunta':
            with transaction.atomic():
                try:
                    id_periodo = int(encrypt(request.POST['id']))
                    form = PreguntaPeriodoPlanificacionForm(request.POST)
                    if form.is_valid() and form.validador(id_periodo):
                        pregunta_pp = PreguntaPeriodoPlanificacion(periodo_id=id_periodo,
                                                                  pregunta=form.cleaned_data['pregunta'])
                        pregunta_pp.save(request)
                        diccionario = {'id': pregunta_pp.id,
                                       'id_periodo': id_periodo,
                                       'pregunta': pregunta_pp.pregunta,
                                       'requerido': pregunta_pp.requerido
                                       }
                        log(u'Agrego pregunta de periodo planificacion: %s' % pregunta_pp, request, "add")
                        return JsonResponse({'result': True, 'data_return': True, 'mensaje': u'Guardado con exito', 'data': diccionario})
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()],
                                             "mensaje": "Error en el formulario"})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "mensaje": 'Error: {}'.format(ex)}, safe=False)

        if action == 'editpregunta':
            with transaction.atomic():
                try:
                    pregunta = PreguntaPeriodoPlanificacion.objects.get(pk=int(request.POST['id']))
                    if request.POST['name'] == 'requerido':
                        pregunta.requerido = eval(request.POST['val'].capitalize())
                        pregunta.save(request)
                        log(u'Edito estado mostrar de requisito : %s' % (pregunta), request,
                            "editrequisito")

                    if request.POST['name'] == 'pregunta':
                        pregunta.pregunta = eval(request.POST['val'].capitalize())
                        pregunta.save(request)
                        log(u'Edito estado opcional de requisito : %s' % (pregunta), request,
                            "editrequisito")

                    return JsonResponse({"result": True, 'mensaje': 'Cambios guardados'})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': False, "mensaje": 'Error: {}'.format(ex)}, safe=False)

        if action == 'delpregunta':
            with transaction.atomic():
                try:
                    instancia = PreguntaPeriodoPlanificacion.objects.get(pk=int(request.POST['id']))
                    instancia.status = False
                    instancia.save(request)
                    log(u'Elimino pregunta de la planificacion: %s' % instancia, request, "del")
                    res_json = {"error": False, "mensaje": 'Registro eliminado'}
                except Exception as ex:
                    res_json = {'error': True, "message": "Error: {}".format(ex)}
                return JsonResponse(res_json, safe=False)

        if action == 'validarpartida':
            try:
                form = ValidarPartidaPlanificacionForm(request.POST)
                partida = Partida.objects.get(id=int(encrypt(request.POST['id'])))
                if form.is_valid():
                    historial = HistorialPartida(partida=partida,
                                                                estado=form.cleaned_data['estado'],
                                                                persona=persona,
                                                                observacion=form.cleaned_data['observacion'])
                    historial.save(request)
                    partida.estado=historial.estado
                    partida.save()
                    log(u'Validó partida: %s' % historial, request, "add")
                    return JsonResponse({'result': False, 'mensaje': u'Guardado con exito'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': u'Error al guardar los datos'})


        return JsonResponse({"result": "bad", "mensaje": u"Solicitud Incorrecta."})
    else:
        # fin get
        adduserdata(request, data)
        if 'action' in request.GET:
            data['action'] = action = request.GET['action']

            if action == 'addconvocatoria':
                try:
                    data['title'] = u'Adiccionar Convocatoria'
                    form = ConvocatoriaForm()
                    form.fields['fechainicio'].widget = CustomDateInput(attrs={'type': 'date', 'class': 'form-control', 'formwidth': '50%'})
                    form.fields['fechafin'].widget = CustomDateInput(attrs={'type': 'date', 'class': 'form-control', 'formwidth': '50%'})
                    data['form'] = form
                    template = get_template("postulate/adm_postulate/modal/formconvocatoria.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            if action == 'editconvocatoria':
                try:
                    data['title'] = u'Editar Convocatoria'
                    data['convocatoria'] = convocatoria = Convocatoria.objects.get(id=int(encrypt(request.GET['id'])))
                    initial = model_to_dict(convocatoria)
                    form = ConvocatoriaForm(initial=initial)
                    form.fields['fechainicio'].widget = CustomDateInput(attrs={'type': 'date', 'class': 'form-control', 'formwidth': '50%'})
                    form.fields['fechafin'].widget = CustomDateInput(attrs={'type': 'date', 'class': 'form-control', 'formwidth': '50%'})
                    form.fields['fechainicio'].initial = str(convocatoria.finicio)
                    form.fields['fechafin'].initial = str(convocatoria.ffin)
                    data['form'] = form
                    template = get_template("postulate/adm_postulate/modal/formconvocatoria.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            if action == 'listarperiodoplanificacion':
                try:
                    data['title'] = u'Periodos de planificación'
                    listado = PeriodoPlanificacion.objects.filter(status=True)
                    paging = MiPaginador(listado, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['listado'] = page.object_list
                    data['list_count'] = len(listado)
                    return render(request, "postulate/adm_postulate/listarperiodosplanificacion.html", data)
                except Exception as ex:
                    pass

            if action == 'listarpartidasplanificacion':
                try:
                    data['title'] = u'Partidas'
                    idpplanificacion = int(encrypt(request.GET['id']))
                    data['periodoplanificacion'] = periodoplanificacion = PeriodoPlanificacion.objects.get(id=idpplanificacion)
                    search, filtro, url_vars = request.GET.get('s', ''), (Q(status=True) & Q(periodoplanificacion=periodoplanificacion)), ''
                    url_vars = '&id={}&action={}'.format(request.GET['id'], action)
                    if search:
                        data['search'] = search
                        url_vars += "&search={}".format(search)
                        filtro = filtro & (Q(denominacionpuesto__descripcion__icontains=search) | Q(codpartida__icontains=search) | Q(estado__icontains=search))
                    listado = Partida.objects.filter(filtro).order_by('-id')
                    paging = MiPaginador(listado, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data["url_vars"] = url_vars
                    data['listado'] = page.object_list
                    data['list_count'] = len(listado)
                    return render(request, "postulate/adm_periodoplanificacion/listarpartidasplanificacion.html", data)
                except Exception as ex:
                    pass

            if action == 'listarpartidas':
                try:
                    data['title'] = u'Partidas'
                    idconvocatoria = int(encrypt(request.GET['id']))
                    data['convocatoria'] = convocatoria = Convocatoria.objects.get(id=idconvocatoria)
                    search, filtro, url_vars = request.GET.get('s', ''), (Q(status=True) & Q(convocatoria=convocatoria)), ''
                    url_vars = '&id={}&action={}'.format(request.GET['id'], action)
                    if search:
                        data['search'] = search
                        url_vars += "&search={}".format(search)
                        filtro = filtro & (Q(denominacionpuesto__descripcion__icontains=search) | Q(codpartida__icontains=search))

                    listado = Partida.objects.filter(filtro).order_by('-id')
                    paging = MiPaginador(listado, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data["url_vars"] = url_vars
                    data['listado'] = page.object_list
                    data['list_count'] = len(listado)
                    return render(request, "postulate/adm_postulate/listarpartidas.html", data)
                except Exception as ex:
                    pass

            if action == 'buscarasignaturas':
                try:
                    # id = request.GET['id']
                    q = request.GET['q'].upper().strip()
                    s = q.split(" ")
                    querybase = Asignatura.objects.filter(status=True)
                    # if len(s) == 1:
                    per = querybase.filter((Q(nombre__icontains=q) | Q(codigo__icontains=q))).distinct()[:30]
                    # elif len(s) == 2:
                    #     per = querybase.filter((Q(nombre__contains=s[0]) & Q(nombre__contains=s[1])) | (Q(abreviatura__icontains=s[0]) & Q(abreviatura__icontains=s[1]))).filter(status=True).distinct()[:30]
                    # else:
                    #     per = querybase.filter((Q(nombre__contains=s[0]) & Q(nombre__contains=s[1])) | (Q(abreviatura__contains=s[0]) & Q(abreviatura__contains=s[1]))).filter(status=True).distinct()[:30]
                    data = {"result": "ok", "results": [{"id": x.id, "name": "{} - {}".format(x.codigo, x.nombre)} for x in per]}
                    return JsonResponse(data)
                except Exception as ex:
                    pass

            if action == 'buscartitulos_old':
                try:
                    # id = request.GET['id']
                    q = request.GET['q'].upper().strip()
                    s = q.split(" ")
                    querybase = Titulo.objects.filter(nivel_id__in=[3, 4, 21, 22, 23], status=True)
                    if len(s) == 1:
                        per = querybase.filter((Q(nombre__icontains=q) | Q(abreviatura__icontains=q)), Q(status=True)).distinct()[:30]
                    elif len(s) == 2:
                        per = querybase.filter((Q(nombre__contains=s[0]) & Q(nombre__contains=s[1])) | (Q(abreviatura__icontains=s[0]) & Q(abreviatura__icontains=s[1]))).filter(status=True).distinct()[:30]
                    else:
                        per = querybase.filter((Q(nombre__contains=s[0]) & Q(nombre__contains=s[1])) | (Q(abreviatura__contains=s[0]) & Q(abreviatura__contains=s[1]))).filter(status=True).distinct()[:30]
                    data = {"result": "ok", "results": [{"id": x.id, "name": "{} - {}".format(x.abreviatura, x.nombre)} for x in per]}
                    return JsonResponse(data)
                except Exception as ex:
                    pass

            if action == 'buscartitulos':
                try:
                    # id = request.GET['id']
                    q = request.GET['q'].upper().strip()
                    s = q.split(" ")
                    querybase = Titulo.objects.filter(nivel_id__in=[3, 4, 21, 22, 23,30], status=True)
                    per = querybase.filter((Q(nombre__icontains=q) | Q(abreviatura__icontains=q)), Q(status=True)).distinct()[:100]
                    data = {"result": "ok", "results": [{"id": x.id, "name": "{} - {}".format(x.abreviatura, x.nombre)} for x in per]}
                    return JsonResponse(data)
                except Exception as ex:
                    pass

            if action == 'listcampoespecifico':
                try:
                    campoamplio = request.GET.get('campoamplio')
                    listcampoamplio = campoamplio
                    if len(campoamplio) > 1:
                        listcampoamplio = campoamplio.split(',')
                    querybase = SubAreaConocimientoTitulacion.objects.filter(status=True, areaconocimiento__in=listcampoamplio).order_by('codigo')
                    if 'q' in request.GET:
                        q = request.GET['q'].upper().strip()
                        if q != 'UNDEFINED':
                            querybase = querybase.filter((Q(nombre__icontains=q) | Q(codigo__icontains=q))).distinct()[:30]
                    data = {"result": "ok", "results": [{"id": x.id, "idca": x.areaconocimiento.id, "name": "{} - {}".format(x.codigo, x.nombre)} for x in querybase]}
                    return JsonResponse(data)
                except Exception as ex:
                    pass

            if action == 'listcampodetallado':
                try:
                    campoespecifico = request.GET.get('campoespecifico')
                    listcampoespecifico = campoespecifico
                    if len(campoespecifico) > 1:
                        listcampoespecifico = campoespecifico.split(',')
                    querybase = SubAreaEspecificaConocimientoTitulacion.objects.filter(status=True, areaconocimiento__in=listcampoespecifico).order_by('codigo')
                    if 'q' in request.GET:
                        q = request.GET['q'].upper().strip()
                        if q != 'UNDEFINED':
                            querybase = querybase.filter((Q(nombre__icontains=q) | Q(codigo__icontains=q))).distinct()[:30]
                    data = {"result": "ok", "results": [{"id": x.id, "name": "{} - {}".format(x.codigo, x.nombre)} for x in querybase]}
                    return JsonResponse(data)
                except Exception as ex:
                    pass


            if action == 'addperiodoplanificacion':
                try:
                    data['title'] = u'Adicionar Periodo Planificación'
                    form = PeriodoPlanificacionForm()
                    form.fields['fechainicio'].widget = CustomDateInput(attrs={'type': 'date', 'class': 'form-control', 'formwidth': '50%'})
                    form.fields['fechafin'].widget = CustomDateInput(attrs={'type': 'date', 'class': 'form-control', 'formwidth': '50%'})
                    data['form'] = form
                    template = get_template("postulate/adm_postulate/modal/formperiodoplanificacion.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            if action == 'editperiodoplanificacion':
                try:
                    data['title'] = u'Editar Periodo de planificacion'
                    data['pplanificacion'] = pplanificacion = PeriodoPlanificacion.objects.get(id=int(encrypt(request.GET['id'])))
                    initial = model_to_dict(pplanificacion)
                    form = PeriodoPlanificacionForm(initial=initial)
                    form.fields['fechainicio'].widget = CustomDateInput(attrs={'type': 'date', 'class': 'form-control', 'formwidth': '50%'})
                    form.fields['fechafin'].widget = CustomDateInput(attrs={'type': 'date', 'class': 'form-control', 'formwidth': '50%'})
                    form.fields['fechainicio'].initial = str(pplanificacion.finicio)
                    form.fields['fechafin'].initial = str(pplanificacion.ffin)
                    data['form'] = form
                    template = get_template("postulate/adm_postulate/modal/formperiodoplanificacion.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            if action == 'addpartida':
                try:
                    data['title'] = u'Adiccionar Partida'
                    form = PartidaForm()
                    form.fields['asignatura'].queryset = Asignatura.objects.none()
                    form.fields['titulos'].queryset = Titulo.objects.none()
                    data['form'] = form
                    data['idc'] = request.GET['idc']
                    return render(request, "postulate/adm_postulate/formpartida.html", data)
                except Exception as ex:
                    pass

            if action == 'editpartida':
                try:
                    data['title'] = u'Editar Partida'
                    data['partida'] = partida = Partida.objects.get(id=int(encrypt(request.GET['id'])))
                    data['partidaasignatura'] = partidaasignatura = PartidaAsignaturas.objects.filter(partida=partida, status=True).values_list('asignatura_id')
                    asignatura = Asignatura.objects.filter(id__in=partidaasignatura)
                    data['idc'] = request.GET['idc']
                    form = PartidaForm(initial={'codpartida': partida.codpartida,
                                                # 'titulo': partida.titulo,
                                                # 'descripcion': partida.descripcion,
                                                'campoamplio': partida.campoamplio.all(),
                                                'denominacionpuesto': partida.denominacionpuesto,
                                                'campoespecifico': partida.campoespecifico.all(),
                                                'campodetallado': partida.campodetallado.all(),
                                                'carrera': partida.carrera,
                                                'titulos': partida.titulos.all(),
                                                'nivel': partida.nivel,
                                                'modalidad': partida.modalidad,
                                                'dedicacion': partida.dedicacion,
                                                'jornada': partida.jornada,
                                                'rmu': partida.rmu,
                                                'vigente': partida.vigente,
                                                'asignatura': asignatura})
                    form.editar(partidaasignatura, partida)
                    data['form'] = form
                    return render(request, "postulate/adm_postulate/formpartida.html", data)
                except Exception as ex:
                    pass

            if action == 'infopartida':
                try:
                    data['partida'] = partida = Partida.objects.get(pk=int(request.GET['id']))
                    data['partidaasignaturas'] = PartidaAsignaturas.objects.filter(partida=partida, status=True)
                    data['detallecapacitacion'] = DetalleCapacitacionPlanificacion.objects.filter(partida=partida, status=True)
                    template = get_template("postulate/adm_periodoplanificacion/modal/formdetallepartida.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

            if action == 'addticompeplanificacion':
                try:
                    data['title'] = u'Adiccionar Tipo Competencia Planificacion'
                    data['action'] = request.GET['action']
                    form = TipoCompetenciaPlanificacionForm()
                    data['form'] = form
                    template = get_template("postulate/adm_periodoplanificacion/modal/formticompeplanificacion.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, 'message': str(ex)})

            if action == 'editticompeplanificacion':
                try:
                    data['title'] = u'Editar Tipo Competencia Planificacion'
                    data['action'] = request.GET['action']
                    data['id'] = id = int(encrypt(request.GET['id']))
                    data['filtro'] = filtro = TipoCompetenciaPlanificacion.objects.get(id=id)
                    initial = model_to_dict(filtro)
                    form = TipoCompetenciaPlanificacionForm(initial=initial)
                    data['form'] = form
                    template = get_template("postulate/adm_periodoplanificacion/modal/formticompeplanificacion.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, 'message': str(ex)})

            if action == 'addticompeespeplanificacion':
                try:
                    data['title'] = u'Adiccionar Tipo Competencia Específica Planificacion'
                    data['action'] = request.GET['action']
                    form = TipoCompetenciaEspecificaPlanificacionForm()
                    data['form'] = form
                    template = get_template("postulate/adm_periodoplanificacion/modal/formticompeplanificacion.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, 'message': str(ex)})

            if action == 'editticompeespeplanificacion':
                try:
                    data['title'] = u'Editar Tipo Competencia Específica Planificacion'
                    data['action'] = request.GET['action']
                    data['id'] = id = int(encrypt(request.GET['id']))
                    data['filtro'] = filtro = TipoCompetenciaEspecificaPlanificacion.objects.get(id=id)
                    initial = model_to_dict(filtro)
                    form = TipoCompetenciaEspecificaPlanificacionForm(initial=initial)
                    data['form'] = form
                    template = get_template("postulate/adm_periodoplanificacion/modal/formticompeplanificacion.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, 'message': str(ex)})

            if action == 'addpartidaplanificacion':
                try:
                    data['title'] = u'Adicionar Partida de periodo de planificacion'
                    form = PartidaPlanificacionAddForm()
                    form.fields['asignatura'].queryset = Asignatura.objects.none()
                    form.fields['titulos'].queryset = Titulo.objects.none()
                    data['form'] = form
                    data['idc'] = request.GET['idc']
                    data['tipocap'] = TipoCompetenciaEspecificaPlanificacion.objects.filter(status=True)
                    data['tiempo'] = TIEMPO_CAPACITACION
                    data['formacap'] = TipoCompetenciaPlanificacion.objects.filter(status=True)
                    data['tipocompetencias'] = TipoCompetenciaLaboral.objects.filter(status=True).order_by('id')
                    data['competencia'] = competencia = CompetenciaLaboral.objects.filter(status=True,tipo__id=1)
                    # data['nivel'] = DetalleCompetenciaLaboral.objects.filter(status=True,competencia=competencia.first() )
                    return render(request, "postulate/adm_periodoplanificacion/formpartidaplanificacion.html", data)
                except Exception as ex:
                    pass

            if action == 'editpartidaplanificacion':
                try:
                    data['title'] = u'Editar Partida'
                    data['partida'] = partida = Partida.objects.get(id=int(encrypt(request.GET['id'])))
                    data['armonizacionpartida'] = armonizacionpartida = partida.obtener_armonizacion()
                    data['partidaasignatura'] = partidaasignatura = PartidaAsignaturas.objects.filter(partida=partida, status=True).values_list('asignatura_id')
                    data['detallecapacitacion'] = detallecapacitacion = DetalleCapacitacionPlanificacion.objects.filter(partida=partida, status=True)
                    data['detallecompetencia'] = detallecompetencia = DetalleCompetenciaPlanificacion.objects.filter(partida=partida, status=True)
                    asignatura = Asignatura.objects.filter(id__in=partidaasignatura)
                    data['idc'] = request.GET['idc']
                    form = PartidaPlanificacionAddForm(initial={'codpartida': partida.codpartida,
                                                # 'titulo': partida.titulo,
                                                # 'descripcion': partida.descripcion,
                                                'denominacionpuesto': partida.denominacionpuesto,
                                                'campoamplio': partida.campoamplio.all(),
                                                'campoespecifico': partida.campoespecifico.all(),
                                                'campodetallado': partida.campodetallado.all(),
                                                'carrera': partida.carrera,
                                                'titulos': partida.titulos.all(),
                                                'nivel': partida.nivel,
                                                'modalidad': partida.modalidad,
                                                'dedicacion': partida.dedicacion,
                                                'jornada': partida.jornada,
                                                'rmu': partida.rmu,
                                                'asignatura': asignatura,
                                                'estado': partida.estado,
                                                'temadisertacion': partida.temadisertacion,
                                                'observacion': partida.observacion,
                                                'armonizacion': armonizacionpartida,
                                                })
                    data['tipocap'] = TipoCompetenciaEspecificaPlanificacion.objects.filter(status=True)
                    data['tiempo'] = TIEMPO_CAPACITACION
                    data['formacap'] = TipoCompetenciaPlanificacion.objects.filter(status=True)
                    data['tipocompetencias'] = TipoCompetenciaLaboral.objects.filter(status=True)
                    data['competencias'] = CompetenciaLaboral.objects.filter(status=True)
                    data['nivel'] = DetalleCompetenciaLaboral.objects.filter(status=True)
                    infoDetalle = []
                    infoDetalle2 = []
                    infoDetalle.append(detallecapacitacion)
                    infoDetalle2.append(detallecompetencia)
                    form.editar(partidaasignatura, partida)
                    data['form'] = form
                    return render(request, "postulate/adm_periodoplanificacion/formpartidaplanificacion.html", data)
                except Exception as ex:
                    pass




            if action == 'agendardisertacion':
                try:
                    data['postulacion'] = postulacion = PersonaAplicarPartida.objects.get(id=int(encrypt(request.GET['id'])))
                    convocatoria = ConvocatoriaPostulante.objects.filter(persona=postulacion).first()
                    if convocatoria:
                        initial = model_to_dict(convocatoria)
                        form = AgendaDisertacionForm(initial=initial)
                        data['convocatoria'] = convocatoria
                    else:
                        form = AgendaDisertacionForm()
                    form.fields['fechaasistencia'].widget = CustomDateInput(attrs={'type': 'date', 'class': 'form-control', 'formwidth': '50%'})
                    form.fields['horasistencia'].widget = CustomDateInput(attrs={'type': 'time', 'class': 'form-control', 'formwidth': '50%'})
                    if convocatoria:
                        form.fields['fechaasistencia'].initial = str(convocatoria.fechaasistencia)
                        form.fields['horasistencia'].initial = str(convocatoria.horasistencia)
                    data['form'] = form
                    template = get_template("postulate/adm_postulate/modal/agendaconvocatoriapostulante.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            if action == 'viewtribunalprimera':
                idpartida = int(encrypt(request.GET['id']))
                data['partida'] = partida = Partida.objects.get(id=idpartida)
                data['action'] = action
                data['listado'] = partida.partidatribunal_set.filter(status=True, tipo=1).order_by('id')
                template = get_template("postulate/adm_postulate/modal/viewtribunal.html")
                return JsonResponse({"result": True, 'data': template.render(data), 'footer': True})

            if action == 'viewtribunalsegunda':
                idpartida = int(encrypt(request.GET['id']))
                data['partida'] = partida = Partida.objects.get(id=idpartida)
                data['action'] = action
                data['listado'] = partida.partidatribunal_set.filter(status=True, tipo=2).order_by('id')
                template = get_template("postulate/adm_postulate/modal/viewtribunalsegunda.html")
                return JsonResponse({"result": True, 'data': template.render(data), 'footer': True})

            if action == 'addtribunal':
                try:
                    idperiodo = int(encrypt(request.GET['id']))
                    data['convocatoria'] = periodoplanificacion = PeriodoPlanificacion.objects.get(id=idperiodo)
                    form = TribunalForm()
                    form.cargar_partidasplanificacion(periodoplanificacion)
                    data['form'] = form
                    data['action'] = action
                    template = get_template("postulate/adm_postulate/modal/formtribunal.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as e:
                    print(e)

            if action == 'buscar_persona':
                data = []
                term = request.GET['term'].upper().strip()
                s = term.split(" ")
                if len(s) == 1:
                    query = Persona.objects.filter(Q(apellido1__icontains=s[0]) | Q(cedula__icontains=s[0])).select_related().values('id', 'apellido1', 'apellido2', 'nombres', 'cedula')
                elif len(s) == 2:
                    query = Persona.objects.filter(Q(apellido1__icontains=s[0]) & Q(apellido2__icontains=s[1]) | Q(apellido1__icontains=s[0]) & Q(nombres__icontains=s[1])).select_related().values('id', 'apellido1', 'apellido2', 'nombres', 'cedula')
                else:
                    query = Persona.objects.filter(Q(apellido1__icontains=s[0]) & Q(apellido2__icontains=s[1]) & Q(nombres__icontains=s[2]) | Q(apellido1__icontains=s[0]) & Q(nombres__icontains=s[1]) | Q(apellido1__icontains=s[0]) & Q(nombres__icontains=s[2])).select_related().values('id', 'apellido1', 'apellido2', 'nombres', 'cedula')
                query = query.order_by('apellido1')
                for a in query[0:10]:
                    result = {'id': a['id'], 'text': str('{} {} {} - {}'.format(a['apellido1'], a['apellido2'], a['nombres'], a['cedula']))}
                    data.append(result)
                return HttpResponse(json.dumps(data), content_type='application/json')

            if action == 'actaconformacion':
                try:
                    if 'fecha' in request.GET:
                        fecha = request.GET['fecha']
                        data['fecha_letra'] = fecha_letra(datetime.strptime(fecha, '%d-%m-%Y').date().__str__())
                    else:
                        fecha = datetime.now().date()
                        data['fecha_letra'] = fecha_letra(fecha.__str__())
                    data['partida'] = partida = Partida.objects.get(id=int(request.GET['idp']))

                    data['firmas'] = PartidaTribunal.objects.filter(partida=partida, status=True, tipo=1,firma=True).order_by('id')

                    return conviert_html_to_pdf(
                        'postulate/actas/actaconformacion.html',
                        {
                            'pagesize': 'a4 landscape',
                            'data': data,
                        }
                    )
                except Exception as ex:
                    messages.error(request, str(ex))

            if action == 'actacalificacionmerito':
                try:
                    data['partida'] = partida = Partida.objects.get(id=int(request.GET['idp']))
                    data['firmas'] = PartidaTribunal.objects.filter(partida=partida, status=True, tipo=1).order_by('id')
                    data['participantes'] = participantes = PersonaAplicarPartida.objects.filter(partida=partida, status=True).order_by('persona__apellido1', 'persona__nombres')
                    data['total'] = len(participantes)
                    return conviert_html_to_pdf(
                        'postulate/actas/actacalificacionmerito.html',
                        {'data': data}
                    )
                except Exception as ex:
                    messages.error(request, str(ex))

            if action == 'actacalificacionmerito2':
                try:
                    data['partida'] = partida = Partida.objects.get(id=int(request.GET['idp']))
                    data['firmas'] = PartidaTribunal.objects.filter(partida=partida, status=True, tipo=1).order_by('id')
                    data['participantes'] = participantes = PersonaAplicarPartida.objects.filter(partida=partida, status=True).order_by('persona__apellido1', 'persona__nombres')
                    data['total'] = len(participantes)
                    return conviert_html_to_pdf(
                        'postulate/actas/actacalificacionmeritodesempate.html',
                        {'data': data}
                    )
                except Exception as ex:
                    messages.error(request, str(ex))

            if action == 'actaentrevistatrib2':
                try:
                    data['partida'] = partida = Partida.objects.get(id=int(request.GET['idp']))
                    data['firmas'] = PartidaTribunal.objects.filter(partida=partida, status=True, tipo=2).order_by('id')
                    data['participantes'] = participantes = partida.personaaplicarpartida_set.filter(status=True, estado__in=[1,4,5]).order_by('-nota_final_entrevista')[:partida.convocatoria.nummejorespuntuados]
                    data['total'] = len(participantes)
                    return conviert_html_to_pdf(
                        'postulate/actas/actaentrevista.html',
                        {'data': data}
                    )
                except Exception as ex:
                    messages.error(request, str(ex))

            if action == 'actapuntajefinaltrib2':
                try:
                    data['partida'] = partida = Partida.objects.get(id=int(request.GET['idp']))
                    data['firmas'] = PartidaTribunal.objects.filter(partida=partida, status=True, tipo=2,firma=True).order_by('id')
                    data['participantes'] = participantes = partida.personaaplicarpartida_set.filter(status=True, estado__in=[1,4,5]).order_by('-nota_final_entrevista')[:partida.convocatoria.nummejorespuntuados]
                    data['total'] = len(participantes)
                    return conviert_html_to_pdf(
                        'postulate/actas/actanotafinal.html',
                        {'data': data}
                    )
                except Exception as ex:
                    messages.error(request, str(ex))

            if action == 'listarterminos':
                try:
                    data['id'] = id = encrypt(request.GET['id'])
                    idconvocatoria = int(encrypt(request.GET['id']))
                    data['title'] = u'Terminos y Condiciones'
                    data['convocatoria'] = convocatoria = Convocatoria.objects.get(id=idconvocatoria)
                    search, filtro, url_vars = request.GET.get('s', ''), (Q(status=True)), ''
                    url_vars = '&action={}'.format(action)
                    if search:
                        data['search'] = search
                        url_vars += "&s={}".format(search)
                        filtro = filtro & (Q(descripcion__icontains=search))

                    if id:
                        data['id'] = id
                        url_vars += "&id={}".format(request.GET['id'])
                        filtro = filtro & (Q(convocatoria_id=id))

                    listado = ConvocatoriaTerminosCondiciones.objects.filter(filtro).order_by('descripcion')
                    paging = MiPaginador(listado, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data["url_vars"] = url_vars
                    data['listado'] = page.object_list
                    data['list_count'] = len(listado)
                    return render(request, "postulate/adm_postulate/listarterminos.html", data)
                except Exception as ex:
                    pass

            if action == 'addtermino':
                try:
                    data['termino'] = Convocatoria.objects.get(id=int(encrypt(request.GET['id'])))
                    data['title'] = u'Adicionar Terminos y Condiciones'
                    data['form'] = ConvocatoriaTerminosForm()
                    template = get_template("postulate/adm_postulate/modal/formterminoscondiciones.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            if action == 'edittermino':
                try:
                    data['title'] = u'Editar Terminos y Condiciones'
                    data['termino'] = termino = ConvocatoriaTerminosCondiciones.objects.get(id=int(encrypt(request.GET['id'])))
                    initial = model_to_dict(termino)
                    data['form'] = ConvocatoriaTerminosForm(initial=initial)
                    template = get_template("postulate/adm_postulate/modal/formterminoscondiciones.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            if action == 'exportarpartidas':
                try:
                    pplanificacion = PeriodoPlanificacion.objects.get(id=request.GET['id'])
                    response = HttpResponse(content_type='application/ms-excel')
                    nom_pplanificacion = pplanificacion.nombre.replace(' ', '_').lower()
                    response[
                        'Content-Disposition'] = 'attachment; filename="partidas_periodoplanificacion_{}.xls"'.format(
                        nom_pplanificacion)
                    title = easyxf(
                        'font: name Calibri, color-index black, bold on , height 350; alignment: horiz centre')
                    title2 = easyxf(
                        'font: name Calibri, color-index black, bold on , height 250; alignment: horiz centre')
                    title3 = easyxf(
                        'font: name Calibri, color-index black, bold on , height 250; pattern: pattern solid, fore_colour gray25; alignment: horiz centre; borders: left thin, right thin, top thin, bottom thin')
                    font_style = xlwt.XFStyle()
                    font_style.font.bold = True
                    fuentecabecera = easyxf(
                        'font: name Calibri, color-index black, bold on; pattern: pattern solid, fore_colour gray25; alignment: horiz centre; borders: left thin, right thin, top thin, bottom thin')
                    fuentenormal = easyxf(
                        'font: name Calibri, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
                    style2 = easyxf('borders: left thin, right thin, top thin, bottom thin; alignment: horiz left')
                    wb = xlwt.Workbook(encoding='utf-8')
                    wc = openxl.Workbook()
                    ws = wb.add_sheet('partidas')
                    wa = wc.active
                    ws.write_merge(0, 0, 0, 8, 'UNIVERSIDAD ESTATAL ESTATAL DE MILAGRO', title)
                    ws.write_merge(1, 1, 0, 8, '{}'.format(pplanificacion.nombre), title2)
                    ws.write_merge(2, 2, 4, 6, 'INSTRUCCION FORMAL', title3)
                    ws.write_merge(2, 2, 7, 9, 'DESEMPATE', title3)
                    ws.write_merge(2, 2, 10, 13, 'COMPETENCIAS', title3)
                    row_num = 3
                    columns = [
                        ('COD. PARTIDA', 10000),
                        # ('TITULO', 10000),
                        ('DENOMINACION PUESTO', 10000),
                        # ('DESCRIPCION', 10000),
                        ('CAMPO AMPLIO', 10000),
                        ('CAMPO ESPECIFICO', 10000),
                        ('CAMPO DETALLADO', 10000),
                        ('MAESTRIAS', 30000),
                        ('PHD', 30000),
                        ('TERCER NIVEL', 30000),
                        ('COMPETENCIA 1', 30000),
                        ('COMPETENCIA 2', 30000),
                        ('COMPETENCIA 3', 30000),
                        ('COMPETENCIA 4', 30000),
                        ('', 30000),
                        ('FACULTAD', 10000),
                        ('CARRERA', 10000),
                        ('MODALIDAD', 10000),
                        ('JORNADA', 10000),
                        ('CARGA HORARIA', 10000),
                        ('REMUNERACIÓN MENSUAL UNIFICADA', 4000),
                        ('VIGENTE', 4000),
                        ('TOTAL POSTULANTES', 10000),
                        ('TEMA DISERTACIÓN', 4000),
                        ('OBSERVACIÓN', 10000), ]
                    listado = Partida.objects.filter(status=True, periodoplanificacion=pplanificacion).order_by('id')
                    partidas = listado.annotate(
                        total_asig=Count('partidaasignaturas__id', filter=Q(partidaasignaturas__status=True)))
                    cant_max_asig = partidas.aggregate(total=Max('total_asig'))['total']
                    if not cant_max_asig:
                        cant_max_asig=1
                    column_max=29+cant_max_asig
                    ws.write_merge(2, 2, 30, column_max, 'PAQUETE', title3)
                    for i in range(1, cant_max_asig+1):
                        columns.append((f'ASIGNATURA { i }', 17000))
                    font_style = xlwt.XFStyle()
                    font_style.font.bold = True
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], fuentecabecera)
                        ws.col(col_num).width = columns[col_num][1]
                    row_num += 1


                    for det in listado:
                        detPartida = DetalleCapacitacionPlanificacion.objects.filter(status=True,
                                                                                     partida=det.id).order_by('id')
                        ws.write(row_num, 0, det.codpartida, style2)
                        # ws.write(row_num, 1, det.titulo, style2)
                        ws.write(row_num, 1, det.denominacionpuesto.__str__() if det.denominacionpuesto else '', style2)
                        # ws.write(row_num, 3, det.descripcion, style2)
                        campo_amplio = ''
                        for ca in det.campoamplio.all():
                            campo_amplio += '{}, '.format(ca.__str__())
                        ws.write(row_num, 2, campo_amplio, style2)
                        campo_especifico = ''
                        for ca in det.campoespecifico.all():
                            campo_especifico += '{}, '.format(ca.__str__())
                        ws.write(row_num, 3, campo_especifico, style2)
                        campo_detallado = ''
                        for ca in det.campodetallado.all():
                            campo_detallado += '{}, '.format(ca.__str__())
                        ws.write(row_num, 4, campo_detallado, style2)
                        titulos_relacionados = ''
                        tercer_nivel = ", ".join([t.nombre for t in det.titulos.filter(nivel_id=3)])
                        cuarto_nivel = ", ".join([t.nombre for t in det.titulos.filter(nivel_id=4).exclude(abreviatura__icontains='ph')])
                        phd_nivel = ", ".join([t.nombre for t in det.titulos.filter(nivel_id=4, abreviatura__icontains='ph')])
                        ws.write(row_num, 5, cuarto_nivel, style2)
                        ws.write(row_num, 6, phd_nivel, style2)
                        ws.write(row_num, 7, tercer_nivel , style2)
                        canttiempo = ''
                        desccapacitacion = ''
                        contaux=8
                        competencias = detPartida.filter(status=True).distinct('tipocompetencia_id').order_by('tipocompetencia_id')
                        for competencia in competencias:
                            competenciastr = ''
                            for e in detPartida.filter(tipocompetencia=competencia.tipocompetencia,status=True).order_by('tipocompetencia_id','id'):
                                competenciastr = competenciastr +'[%s %s] %s %s (%s) \n' %(e.canttiempocapacitacion,e.get_tiempocapacitacion_display(),e.tipocompetencia.nombre,e.cespecifica.nombre,e.descripcioncapacitacion )
                            ws.write(row_num, contaux, competenciastr, style2)
                            contaux+=1
                        contaux=12
                        if det.carrera:
                            if det.carrera.mi_coordinacion():
                                ws.write(row_num, contaux+1, det.carrera.mi_coordinacion(), style2)
                            else:
                                ws.write(row_num, contaux+1, '', style2)
                        ws.write(row_num, contaux+2, det.carrera.__str__() if det.carrera else '', style2)
                        ws.write(row_num, contaux+3, det.get_modalidad_display(), style2)
                        ws.write(row_num, contaux+4, det.get_jornada_display(), style2)
                        ws.write(row_num, contaux+5, det.get_dedicacion_display(), style2)
                        ws.write(row_num, contaux+6, det.rmu, style2)
                        ws.write(row_num, contaux+7, 'SI' if det.vigente else 'NO', style2)
                        ws.write(row_num, contaux+8, det.total_postulantes(), style2)
                        ws.write(row_num, contaux+9, det.temadisertacion, style2)
                        ws.write(row_num, contaux+10, det.observacion, style2)
                        col_num = contaux+10
                        asignaturas = det.partidas_asignaturas()
                        cant = asignaturas.count()
                        for i in range(0, cant_max_asig):
                            col_num+=1
                            nombre_asignatura=''
                            if i < cant:
                                nombre_asignatura = det.partidas_asignaturas()[i].asignatura.nombre
                            ws.write(row_num, col_num, nombre_asignatura, style2)

                        row_num += 1
                    wb.save(response)
                    return response
                except Exception as ex:
                    textoerror = '{} Linea:{}'.format(ex, sys.exc_info()[-1].tb_lineno)
                    print(textoerror)
                    messages.success(request, str(ex))

            if action == 'excel_postulantes__all':
                try:
                    response = HttpResponse(content_type='application/ms-excel')
                    response['Content-Disposition'] = 'attachment; filename="postulantes_all.xls"'
                    title = easyxf('font: name Calibri, color-index black, bold on , height 350; alignment: horiz centre')
                    title2 = easyxf('font: name Calibri, color-index black, bold on , height 250; alignment: horiz centre')
                    font_style = xlwt.XFStyle()
                    font_style.font.bold = True
                    fuentecabecera = easyxf('font: name Calibri, color-index black, bold on; pattern: pattern solid, fore_colour gray25; alignment: horiz centre; borders: left thin, right thin, top thin, bottom thin')
                    fuentenormal = easyxf('font: name Calibri, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
                    style2 = easyxf('borders: left thin, right thin, top thin, bottom thin; alignment: horiz centre')
                    wb = xlwt.Workbook(encoding='utf-8')
                    ws = wb.add_sheet('postulantes')
                    ws.write_merge(0, 0, 0, 8, 'UNIVERSIDAD ESTATAL ESTATAL DE MILAGRO', title)
                    ws.write_merge(1, 1, 0, 8, 'LISTADO DE POSTULANTES', title2)
                    row_num = 2
                    columns = [
                        ('Convocatoria', 10000),
                        ('Categoria', 10000),
                        ('Partida', 30000),
                        ('Carrera', 30000),
                        ('Asignaturas', 30000),
                        ('Dedicación', 10000),
                        ('RMU', 10000),
                        ('Nivel', 10000),
                        ('Modalidad', 10000),
                        ('Jornada', 10000),
                        ('Campo Amplio', 30000),
                        ('Campo Especifico', 30000),
                        ('Campo Detallado', 30000),
                        ('Cod. Unico', 10000),
                        ('Apellidos', 10000),
                        ('Nombres', 10000),
                        ('Titulo', 25000),
                        ('Archivo', 10000),
                        ('Identificación', 10000),
                        ('Correo', 10000),
                        ('Telf.', 10000),
                        ('Telf. Conv.', 10000),
                        ('Estado Calificación', 10000),
                        ('¿Tiene Formación?', 10000),
                        ('¿Tiene Experiencia?', 10000),
                        ('¿Tiene Capacitación?', 10000),
                        ('¿Tiene Publicaciones?', 10000),
                        ('¿Tiene Certificación Idiomas?', 10000),
                        ('Nota Formación', 10000),
                        ('Obs. Formación', 20000),
                        ('Nota Exp. Docente', 10000),
                        ('Obs. Exp. Docente', 20000),
                        ('Nota Exp. Administrativo', 10000),
                        ('Obs. Exp. Administrativo', 20000),
                        ('Nota Capacitación', 10000),
                        ('Obs. Capacitación', 20000),
                        ('¿Aplico Desempate?', 10000),
                        ('Nota Desempate', 10000),
                        ('Nota Final', 10000),
                        ('Obs. Final', 20000),
                        ('Estado', 10000),
                        ('¿Calificada?', 10000),
                        ('¿Apelo?', 10000),
                        ('Estado Apelación', 10000),
                        ('Obs. Apelación', 20000),
                        ('Usuario Revisión Meritos', 10000),
                        ('Fecha Revisión Meritos', 10000),
                        ('Usuario Revisión Apelación', 10000),
                        ('Fecha Revisión Apelación', 10000),
                        ('Usuario Revisión Desempate', 10000),
                        ('Fecha Revisión Desempate', 10000),
                    ]
                    font_style = xlwt.XFStyle()
                    font_style.font.bold = True
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], fuentecabecera)
                        ws.col(col_num).width = columns[col_num][1]
                    row_num += 1
                    convocatoria = Convocatoria.objects.get(status=True,pk = int(request.GET['id']))
                    listado = PersonaAplicarPartida.objects.filter(status=True, partida__convocatoria=convocatoria).order_by('partida__convocatoria__descripcion')
                    for det in listado:
                        if det.persona.mis_titulaciones().count() > 0:
                            row_count = row_num + det.persona.mis_titulaciones().count() - 1
                            titulos = det.persona.mis_titulaciones()
                            ws.write_merge(row_num, row_count, 0, 0, det.partida.convocatoria.descripcion, style2)
                            ws.write_merge(row_num, row_count, 1, 1, det.partida.convocatoria.tipocontrato.nombre if det.partida.convocatoria.tipocontrato else '', style2)
                            ws.write_merge(row_num, row_count, 2, 2, str(det.partida), style2)
                            ws.write_merge(row_num, row_count, 3, 3, det.partida.carrera.__str__() if det.partida.carrera else '', style2)
                            asignaturas_ = ''
                            for asig in det.partida.partidas_asignaturas():
                                asignaturas_ += '{}, '.format(asig.__str__())
                            ws.write_merge(row_num, row_count, 4, 4, asignaturas_, style2)
                            ws.write_merge(row_num, row_count, 5, 5, det.partida.get_dedicacion_display(), style2)
                            ws.write_merge(row_num, row_count, 6, 6, det.partida.rmu, style2)
                            ws.write_merge(row_num, row_count, 7, 7, det.partida.get_nivel_display(), style2)
                            ws.write_merge(row_num, row_count, 8, 8, det.partida.get_modalidad_display(), style2)
                            ws.write_merge(row_num, row_count, 9, 9, det.partida.get_jornada_display(), style2)
                            campo_amplio = ''
                            for ca in det.partida.campoamplio.all():
                                campo_amplio += '{}, '.format(ca.__str__())
                            ws.write_merge(row_num, row_count, 10, 10, campo_amplio, style2)
                            campo_especifico = ''
                            for ca in det.partida.campoespecifico.all():
                                campo_especifico += '{}, '.format(ca.__str__())
                            ws.write_merge(row_num, row_count, 11, 11, campo_especifico, style2)
                            campo_detallado = ''
                            for ca in det.partida.campodetallado.all():
                                campo_detallado += '{}, '.format(ca.__str__())
                            ws.write_merge(row_num, row_count, 12, 12, campo_detallado, style2)
                            ws.write_merge(row_num, row_count, 13, 13, det.pk, style2)
                            ws.write_merge(row_num, row_count, 14, 14, "{} {}".format(det.persona.apellido1, det.persona.apellido2), style2)
                            ws.write_merge(row_num, row_count, 15, 15, "{}".format(det.persona.nombres), style2)

                            ws.write_merge(row_num, row_count, 18, 18, det.persona.cedula, style2)
                            ws.write_merge(row_num, row_count, 19, 19, det.persona.email, style2)
                            ws.write_merge(row_num, row_count, 20, 20, det.persona.telefono, style2)
                            ws.write_merge(row_num, row_count, 21, 21, det.persona.telefono_conv, style2)
                            ws.write_merge(row_num, row_count, 22, 22, det.get_estado_display(), style2)
                            ws.write_merge(row_num, row_count, 23, 23, 'SI' if det.tiene_formacionacademica() else 'NO', style2)
                            ws.write_merge(row_num, row_count, 24, 24, 'SI' if det.tiene_experienciapartida() else 'NO', style2)
                            ws.write_merge(row_num, row_count, 25, 25, 'SI' if det.tiene_capacitaciones() else 'NO', style2)
                            ws.write_merge(row_num, row_count, 26, 26, 'SI' if det.tiene_publicaciones() else 'NO', style2)
                            ws.write_merge(row_num, row_count, 27, 27, 'SI' if det.tiene_idiomas() else 'NO', style2)
                            ws.write_merge(row_num, row_count, 28, 28, det.pgradoacademico, style2)
                            ws.write_merge(row_num, row_count, 29, 29, det.obsgradoacademico, style2)
                            ws.write_merge(row_num, row_count, 30, 30, det.pexpdocente, style2)
                            ws.write_merge(row_num, row_count, 31, 31, det.obsexperienciadoc, style2)
                            ws.write_merge(row_num, row_count, 32, 32, det.pexpadministrativa, style2)
                            ws.write_merge(row_num, row_count, 33, 33, det.obsexperienciaadmin, style2)
                            ws.write_merge(row_num, row_count, 34, 34, det.pcapacitacion, style2)
                            ws.write_merge(row_num, row_count, 35, 35, det.obscapacitacion, style2)
                            ws.write_merge(row_num, row_count, 36, 36, 'SI' if det.aplico_desempate else 'NO', style2)
                            ws.write_merge(row_num, row_count, 37, 37, det.nota_desempate, style2)
                            ws.write_merge(row_num, row_count, 38, 38, det.nota_final_meritos, style2)
                            ws.write_merge(row_num, row_count, 39, 39, det.obsgeneral, style2)
                            ws.write_merge(row_num, row_count, 40, 40, det.get_estado_display(), style2)
                            ws.write_merge(row_num, row_count, 41, 41, 'SI' if det.calificada else 'NO', style2)
                            ws.write_merge(row_num, row_count, 42, 42, 'SI' if det.solapelacion else 'NO', style2)
                            ws.write_merge(row_num, row_count, 43, 43, det.traer_apelacion().get_estado_display() if det.traer_apelacion() else '', style2)
                            obs_revisor, revisado_por, fecha_revision = '', '', ''
                            if det.traer_apelacion():
                                if det.traer_apelacion().estado != 0:
                                    obs_revisor = det.traer_apelacion().observacion_revisor
                                    revisado_por = det.traer_apelacion().revisado_por.username if det.traer_apelacion().revisado_por else ''
                                    fecha_revision = str(det.traer_apelacion().fecha_revision)
                            ws.write_merge(row_num, row_count, 44, 44, obs_revisor, style2)
                            ws.write_merge(row_num, row_count, 45, 45, revisado_por, style2)
                            ws.write_merge(row_num, row_count, 46, 46, fecha_revision, style2)
                            ws.write_merge(row_num, row_count, 47, 47, det.revisado_por.username if det.revisado_por else '', style2)
                            ws.write_merge(row_num, row_count, 48, 48, str(det.fecha_revision) if det.fecha_revision else '', style2)
                            ws.write_merge(row_num, row_count, 49, 49, det.desempate_revisado_por.username if det.desempate_revisado_por else '', style2)
                            ws.write_merge(row_num, row_count, 50, 50, str(det.desempate_fecha_revision) if det.desempate_fecha_revision else '', style2)
                            for dato in titulos:
                                ws.write(row_num, 16, "{}".format(dato.titulo), style2)
                                ws.write(row_num, 17, "https://sga.unemi.edu.ec/{}".format(dato.archivo.url) if dato.archivo else '', style2)
                                row_num += 1
                        else:
                            ws.write(row_num, 0, det.partida.convocatoria.descripcion, style2)
                            ws.write(row_num, 1, det.partida.convocatoria.tipocontrato.nombre if det.partida.convocatoria.tipocontrato else '', style2)
                            ws.write(row_num, 2, str(det.partida), style2)
                            ws.write(row_num, 3, det.partida.carrera.__str__() if det.partida.carrera else '', style2)
                            asignaturas_ = ''
                            for asig in det.partida.partidas_asignaturas():
                                asignaturas_ += '{}, '.format(asig.__str__())
                            ws.write(row_num, 4, asignaturas_, style2)
                            ws.write(row_num, 5, det.partida.get_dedicacion_display(), style2)
                            ws.write(row_num, 6, det.partida.rmu, style2)
                            ws.write(row_num, 7, det.partida.get_nivel_display(), style2)
                            ws.write(row_num, 8, det.partida.get_modalidad_display(), style2)
                            ws.write(row_num, 9, det.partida.get_jornada_display(), style2)
                            campo_amplio = ''
                            for ca in det.partida.campoamplio.all():
                                campo_amplio += '{}, '.format(ca.__str__())
                            ws.write(row_num, 10, campo_amplio, style2)
                            campo_especifico = ''
                            for ca in det.partida.campoespecifico.all():
                                campo_especifico += '{}, '.format(ca.__str__())
                            ws.write(row_num, 11, campo_especifico, style2)
                            campo_detallado = ''
                            for ca in det.partida.campodetallado.all():
                                campo_detallado += '{}, '.format(ca.__str__())
                            ws.write(row_num, 12, campo_detallado, style2)
                            ws.write(row_num, 13, det.pk, style2)
                            ws.write(row_num, 14, "{} {}".format(det.persona.apellido1, det.persona.apellido2), style2)
                            ws.write(row_num, 15, "{}".format(det.persona.nombres), style2)
                            ws.write(row_num, 16, "", style2)
                            ws.write(row_num, 17, "", style2)
                            ws.write(row_num, 18, det.persona.cedula, style2)
                            ws.write(row_num, 19, det.persona.email, style2)
                            ws.write(row_num, 20, det.persona.telefono, style2)
                            ws.write(row_num, 21, det.persona.telefono_conv, style2)
                            ws.write(row_num, 22, det.get_estado_display(), style2)
                            ws.write(row_num, 23, 'SI' if det.tiene_formacionacademica() else 'NO', style2)
                            ws.write(row_num, 24, 'SI' if det.tiene_experienciapartida() else 'NO', style2)
                            ws.write(row_num, 25, 'SI' if det.tiene_capacitaciones() else 'NO', style2)
                            ws.write(row_num, 26, 'SI' if det.tiene_publicaciones() else 'NO', style2)
                            ws.write(row_num, 27, 'SI' if det.tiene_idiomas() else 'NO', style2)
                            ws.write(row_num, 28, det.pgradoacademico, style2)
                            ws.write(row_num, 29, det.obsgradoacademico, style2)
                            ws.write(row_num, 30, det.pexpdocente, style2)
                            ws.write(row_num, 31, det.obsexperienciadoc, style2)
                            ws.write(row_num, 32, det.pexpadministrativa, style2)
                            ws.write(row_num, 33, det.obsexperienciaadmin, style2)
                            ws.write(row_num, 34, det.pcapacitacion, style2)
                            ws.write(row_num, 35, det.obscapacitacion, style2)
                            ws.write(row_num, 36, 'SI' if det.aplico_desempate else 'NO', style2)
                            ws.write(row_num, 37, det.nota_desempate, style2)
                            ws.write(row_num, 38, det.nota_final_meritos, style2)
                            ws.write(row_num, 39, det.obsgeneral, style2)
                            ws.write(row_num, 40, det.get_estado_display(), style2)
                            ws.write(row_num, 41, 'SI' if det.calificada else 'NO', style2)
                            ws.write(row_num, 42, 'SI' if det.solapelacion else 'NO', style2)
                            ws.write(row_num, 43, det.traer_apelacion().get_estado_display() if det.traer_apelacion() else '', style2)
                            obs_revisor, revisado_por, fecha_revision = '', '', ''
                            if det.traer_apelacion():
                                if det.traer_apelacion().estado != 0:
                                    obs_revisor = det.traer_apelacion().observacion_revisor
                                    revisado_por = det.traer_apelacion().revisado_por.username if det.traer_apelacion().revisado_por else ''
                                    fecha_revision = str(det.traer_apelacion().fecha_revision)
                            ws.write(row_num, 44, obs_revisor, style2)
                            ws.write(row_num, 45, revisado_por, style2)
                            ws.write(row_num, 46, fecha_revision, style2)
                            ws.write(row_num, 47, det.revisado_por.username if det.revisado_por else '', style2)
                            ws.write(row_num, 48, str(det.fecha_revision) if det.fecha_revision else '', style2)
                            ws.write(row_num, 49, det.desempate_revisado_por.username if det.desempate_revisado_por else '', style2)
                            ws.write(row_num, 50, str(det.desempate_fecha_revision) if det.desempate_fecha_revision else '', style2)
                            row_num += 1
                    wb.save(response)
                    return response
                except Exception as ex:
                    messages.error(request, str(ex))

            if action == 'excel_postulantes__all_mejores_puntuados':
                try:
                    response = HttpResponse(content_type='application/ms-excel')
                    response['Content-Disposition'] = 'attachment; filename="postulantes_all_mejores_puntuados.xls"'
                    title = easyxf('font: name Calibri, color-index black, bold on , height 350; alignment: horiz centre')
                    title2 = easyxf('font: name Calibri, color-index black, bold on , height 250; alignment: horiz centre')
                    font_style = xlwt.XFStyle()
                    font_style.font.bold = True
                    fuentecabecera = easyxf('font: name Calibri, color-index black, bold on; pattern: pattern solid, fore_colour gray25; alignment: horiz centre; borders: left thin, right thin, top thin, bottom thin')
                    fuentenormal = easyxf('font: name Calibri, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
                    style2 = easyxf('borders: left thin, right thin, top thin, bottom thin; alignment: horiz centre')
                    wb = xlwt.Workbook(encoding='utf-8')
                    ws = wb.add_sheet('postulantes')
                    ws.write_merge(0, 0, 0, 8, 'UNIVERSIDAD ESTATAL ESTATAL DE MILAGRO', title)
                    ws.write_merge(1, 1, 0, 8, 'LISTADO DE POSTULANTES MEJORES PUNTUADOS', title2)
                    row_num = 2
                    columns = [
                        ('Convocatoria', 10000),
                        ('Categoria', 10000),
                        ('Partida', 30000),
                        ('Carrera', 30000),
                        ('Asignaturas', 30000),
                        ('Dedicación', 10000),
                        ('RMU', 10000),
                        ('Nivel', 10000),
                        ('Modalidad', 10000),
                        ('Jornada', 10000),
                        ('Campo Amplio', 30000),
                        ('Campo Especifico', 30000),
                        ('Campo Detallado', 30000),
                        ('Cod. Unico', 10000),
                        ('Apellidos', 10000),
                        ('Nombres', 10000),
                        ('Titulo', 25000),
                        ('Archivo', 10000),
                        ('Identificación', 10000),
                        ('Correo', 10000),
                        ('Telf.', 10000),
                        ('Telf. Conv.', 10000),
                        ('Estado Calificación', 10000),
                        ('¿Tiene Formación?', 10000),
                        ('¿Tiene Experiencia?', 10000),
                        ('¿Tiene Capacitación?', 10000),
                        ('¿Tiene Publicaciones?', 10000),
                        ('¿Tiene Certificación Idiomas?', 10000),
                        ('Nota Formación', 10000),
                        ('Obs. Formación', 20000),
                        ('Nota Exp. Docente', 10000),
                        ('Obs. Exp. Docente', 20000),
                        ('Nota Exp. Administrativo', 10000),
                        ('Obs. Exp. Administrativo', 20000),
                        ('Nota Capacitación', 10000),
                        ('Obs. Capacitación', 20000),
                        ('¿Aplico Desempate?', 10000),
                        ('Nota Desempate', 10000),
                        ('Nota Final', 10000),
                        ('Obs. Final', 20000),
                        ('Estado', 10000),
                        ('¿Calificada?', 10000),
                        ('¿Apelo?', 10000),
                        ('Estado Apelación', 10000),
                        ('Obs. Apelación', 20000),
                        ('Usuario Revisión Meritos', 10000),
                        ('Fecha Revisión Meritos', 10000),
                        ('Usuario Revisión Apelación', 10000),
                        ('Fecha Revisión Apelación', 10000),
                        ('Usuario Revisión Desempate', 10000),
                        ('Fecha Revisión Desempate', 10000),
                    ]
                    font_style = xlwt.XFStyle()
                    font_style.font.bold = True
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], fuentecabecera)
                        ws.col(col_num).width = columns[col_num][1]
                    row_num += 1
                    convocatoria = Convocatoria.objects.get(status=True, pk=int(request.GET['id']))
                    for partidas in Partida.objects.filter(status=True, convocatoria=convocatoria).order_by('convocatoria__descripcion'):
                        listado = partidas.personaaplicarpartida_set.filter(status=True).order_by('-nota_final_meritos')[:3]
                        for det in listado:
                            if det.persona.mis_titulaciones().count() >0:
                                row_count = row_num+det.persona.mis_titulaciones().count()-1
                                titulos = det.persona.mis_titulaciones()
                                ws.write_merge(row_num,row_count, 0,0, det.partida.convocatoria.descripcion, style2)
                                ws.write_merge(row_num,row_count, 1,1, det.partida.convocatoria.tipocontrato.nombre if det.partida.convocatoria.tipocontrato else '', style2)
                                ws.write_merge(row_num,row_count, 2,2, str(det.partida), style2)
                                ws.write_merge(row_num,row_count, 3,3, det.partida.carrera.__str__() if det.partida.carrera else '', style2)
                                asignaturas_ = ''
                                for asig in det.partida.partidas_asignaturas():
                                    asignaturas_ += '{}, '.format(asig.__str__())
                                ws.write_merge(row_num,row_count, 4,4, asignaturas_, style2)
                                ws.write_merge(row_num,row_count, 5,5, det.partida.get_dedicacion_display(), style2)
                                ws.write_merge(row_num,row_count, 6,6, det.partida.rmu, style2)
                                ws.write_merge(row_num,row_count, 7,7, det.partida.get_nivel_display(), style2)
                                ws.write_merge(row_num,row_count, 8,8, det.partida.get_modalidad_display(), style2)
                                ws.write_merge(row_num,row_count, 9,9, det.partida.get_jornada_display(), style2)
                                campo_amplio = ''
                                for ca in det.partida.campoamplio.all():
                                    campo_amplio += '{}, '.format(ca.__str__())
                                ws.write_merge(row_num,row_count ,10,10, campo_amplio, style2)
                                campo_especifico = ''
                                for ca in det.partida.campoespecifico.all():
                                    campo_especifico += '{}, '.format(ca.__str__())
                                ws.write_merge(row_num,row_count, 11,11, campo_especifico, style2)
                                campo_detallado = ''
                                for ca in det.partida.campodetallado.all():
                                    campo_detallado += '{}, '.format(ca.__str__())
                                ws.write_merge(row_num,row_count, 12,12, campo_detallado, style2)
                                ws.write_merge(row_num,row_count, 13,13, det.pk, style2)
                                ws.write_merge(row_num,row_count, 14,14, "{} {}".format(det.persona.apellido1, det.persona.apellido2), style2)
                                ws.write_merge(row_num,row_count, 15,15, "{}".format(det.persona.nombres), style2)

                                ws.write_merge(row_num,row_count, 18,18, det.persona.cedula, style2)
                                ws.write_merge(row_num,row_count, 19,19, det.persona.email, style2)
                                ws.write_merge(row_num,row_count, 20,20, det.persona.telefono, style2)
                                ws.write_merge(row_num,row_count, 21,21, det.persona.telefono_conv, style2)
                                ws.write_merge(row_num,row_count, 22,22, det.get_estado_display(), style2)
                                ws.write_merge(row_num,row_count, 23,23, 'SI' if det.tiene_formacionacademica() else 'NO', style2)
                                ws.write_merge(row_num,row_count, 24,24, 'SI' if det.tiene_experienciapartida() else 'NO', style2)
                                ws.write_merge(row_num,row_count, 25,25, 'SI' if det.tiene_capacitaciones() else 'NO', style2)
                                ws.write_merge(row_num,row_count, 26,26, 'SI' if det.tiene_publicaciones() else 'NO', style2)
                                ws.write_merge(row_num,row_count, 27,27, 'SI' if det.tiene_idiomas() else 'NO', style2)
                                ws.write_merge(row_num,row_count, 28,28, det.pgradoacademico, style2)
                                ws.write_merge(row_num,row_count, 29,29, det.obsgradoacademico, style2)
                                ws.write_merge(row_num,row_count, 30,30, det.pexpdocente, style2)
                                ws.write_merge(row_num,row_count, 31,31, det.obsexperienciadoc, style2)
                                ws.write_merge(row_num,row_count, 32,32, det.pexpadministrativa, style2)
                                ws.write_merge(row_num,row_count, 33,33, det.obsexperienciaadmin, style2)
                                ws.write_merge(row_num,row_count, 34,34, det.pcapacitacion, style2)
                                ws.write_merge(row_num,row_count, 35,35, det.obscapacitacion, style2)
                                ws.write_merge(row_num,row_count, 36,36, 'SI' if det.aplico_desempate else 'NO', style2)
                                ws.write_merge(row_num,row_count, 37,37, det.nota_desempate, style2)
                                ws.write_merge(row_num,row_count, 38,38, det.nota_final_meritos, style2)
                                ws.write_merge(row_num,row_count, 39,39, det.obsgeneral, style2)
                                ws.write_merge(row_num,row_count, 40,40, det.get_estado_display(), style2)
                                ws.write_merge(row_num,row_count, 41,41, 'SI' if det.calificada else 'NO', style2)
                                ws.write_merge(row_num,row_count, 42,42, 'SI' if det.solapelacion else 'NO', style2)
                                ws.write_merge(row_num,row_count, 43,43, det.traer_apelacion().get_estado_display() if det.traer_apelacion() else '', style2)
                                obs_revisor, revisado_por, fecha_revision = '', '', ''
                                if det.traer_apelacion():
                                    if det.traer_apelacion().estado != 0:
                                        obs_revisor = det.traer_apelacion().observacion_revisor
                                        revisado_por = det.traer_apelacion().revisado_por.username if det.traer_apelacion().revisado_por else ''
                                        fecha_revision = str(det.traer_apelacion().fecha_revision)
                                ws.write_merge(row_num,row_count, 44,44, obs_revisor, style2)
                                ws.write_merge(row_num,row_count, 45,45, revisado_por, style2)
                                ws.write_merge(row_num,row_count, 46,46, fecha_revision, style2)
                                ws.write_merge(row_num,row_count, 47,47, det.revisado_por.username if det.revisado_por else '', style2)
                                ws.write_merge(row_num,row_count, 48,48, str(det.fecha_revision) if det.fecha_revision else '', style2)
                                ws.write_merge(row_num,row_count, 49,49, det.desempate_revisado_por.username if det.desempate_revisado_por else '', style2)
                                ws.write_merge(row_num,row_count, 50,50, str(det.desempate_fecha_revision) if det.desempate_fecha_revision else '', style2)
                                for dato in titulos:
                                    ws.write(row_num, 16, "{}".format(dato.titulo), style2)
                                    ws.write(row_num, 17, "https://sga.unemi.edu.ec/{}".format(dato.archivo.url) if dato.archivo else '', style2)
                                    row_num += 1
                            else:
                                ws.write(row_num, 0, det.partida.convocatoria.descripcion, style2)
                                ws.write(row_num, 1, det.partida.convocatoria.tipocontrato.nombre if det.partida.convocatoria.tipocontrato else '', style2)
                                ws.write(row_num, 2, str(det.partida), style2)
                                ws.write(row_num, 3, det.partida.carrera.__str__() if det.partida.carrera else '', style2)
                                asignaturas_ = ''
                                for asig in det.partida.partidas_asignaturas():
                                    asignaturas_ += '{}, '.format(asig.__str__())
                                ws.write(row_num, 4, asignaturas_, style2)
                                ws.write(row_num, 5, det.partida.get_dedicacion_display(), style2)
                                ws.write(row_num, 6, det.partida.rmu, style2)
                                ws.write(row_num, 7, det.partida.get_nivel_display(), style2)
                                ws.write(row_num, 8, det.partida.get_modalidad_display(), style2)
                                ws.write(row_num, 9, det.partida.get_jornada_display(), style2)
                                campo_amplio = ''
                                for ca in det.partida.campoamplio.all():
                                    campo_amplio += '{}, '.format(ca.__str__())
                                ws.write(row_num, 10, campo_amplio, style2)
                                campo_especifico = ''
                                for ca in det.partida.campoespecifico.all():
                                    campo_especifico += '{}, '.format(ca.__str__())
                                ws.write(row_num, 11, campo_especifico, style2)
                                campo_detallado = ''
                                for ca in det.partida.campodetallado.all():
                                    campo_detallado += '{}, '.format(ca.__str__())
                                ws.write(row_num, 12, campo_detallado, style2)
                                ws.write(row_num, 13, det.pk, style2)
                                ws.write(row_num, 14, "{} {}".format(det.persona.apellido1, det.persona.apellido2), style2)
                                ws.write(row_num, 15, "{}".format(det.persona.nombres), style2)
                                ws.write(row_num, 16, "", style2)
                                ws.write(row_num, 17, "", style2)
                                ws.write(row_num, 18, det.persona.cedula, style2)
                                ws.write(row_num, 19, det.persona.email, style2)
                                ws.write(row_num, 20, det.persona.telefono, style2)
                                ws.write(row_num, 21, det.persona.telefono_conv, style2)
                                ws.write(row_num, 22, det.get_estado_display(), style2)
                                ws.write(row_num, 23, 'SI' if det.tiene_formacionacademica() else 'NO', style2)
                                ws.write(row_num, 24, 'SI' if det.tiene_experienciapartida() else 'NO', style2)
                                ws.write(row_num, 25, 'SI' if det.tiene_capacitaciones() else 'NO', style2)
                                ws.write(row_num, 26, 'SI' if det.tiene_publicaciones() else 'NO', style2)
                                ws.write(row_num, 27, 'SI' if det.tiene_idiomas() else 'NO', style2)
                                ws.write(row_num, 28, det.pgradoacademico, style2)
                                ws.write(row_num, 29, det.obsgradoacademico, style2)
                                ws.write(row_num, 30, det.pexpdocente, style2)
                                ws.write(row_num, 31, det.obsexperienciadoc, style2)
                                ws.write(row_num, 32, det.pexpadministrativa, style2)
                                ws.write(row_num, 33, det.obsexperienciaadmin, style2)
                                ws.write(row_num, 34, det.pcapacitacion, style2)
                                ws.write(row_num, 35, det.obscapacitacion, style2)
                                ws.write(row_num, 36, 'SI' if det.aplico_desempate else 'NO', style2)
                                ws.write(row_num, 37, det.nota_desempate, style2)
                                ws.write(row_num, 38, det.nota_final_meritos, style2)
                                ws.write(row_num, 39, det.obsgeneral, style2)
                                ws.write(row_num, 40, det.get_estado_display(), style2)
                                ws.write(row_num, 41, 'SI' if det.calificada else 'NO', style2)
                                ws.write(row_num, 42, 'SI' if det.solapelacion else 'NO', style2)
                                ws.write(row_num, 43, det.traer_apelacion().get_estado_display() if det.traer_apelacion() else '', style2)
                                obs_revisor, revisado_por, fecha_revision = '', '', ''
                                if det.traer_apelacion():
                                    if det.traer_apelacion().estado != 0:
                                        obs_revisor = det.traer_apelacion().observacion_revisor
                                        revisado_por = det.traer_apelacion().revisado_por.username if det.traer_apelacion().revisado_por else ''
                                        fecha_revision = str(det.traer_apelacion().fecha_revision)
                                ws.write(row_num, 44, obs_revisor, style2)
                                ws.write(row_num, 45, revisado_por, style2)
                                ws.write(row_num, 46, fecha_revision, style2)
                                ws.write(row_num, 47, det.revisado_por.username if det.revisado_por else '', style2)
                                ws.write(row_num, 48, str(det.fecha_revision) if det.fecha_revision else '', style2)
                                ws.write(row_num, 49, det.desempate_revisado_por.username if det.desempate_revisado_por else '', style2)
                                ws.write(row_num, 50, str(det.desempate_fecha_revision) if det.desempate_fecha_revision else '', style2)
                                row_num += 1
                    wb.save(response)
                    return response
                except Exception as ex:
                    messages.error(request, str(ex))

            if action == 'excel_postulantes__all_banco_habilitados':
                try:
                    response = HttpResponse(content_type='application/ms-excel')
                    response['Content-Disposition'] = 'attachment; filename="postulantes_banco_habilitantes.xls"'
                    title = easyxf('font: name Calibri, color-index black, bold on , height 350; alignment: horiz centre')
                    title2 = easyxf('font: name Calibri, color-index black, bold on , height 250; alignment: horiz centre')
                    font_style = xlwt.XFStyle()
                    font_style.font.bold = True
                    fuentecabecera = easyxf('font: name Calibri, color-index black, bold on; pattern: pattern solid, fore_colour gray25; alignment: horiz centre; borders: left thin, right thin, top thin, bottom thin')
                    fuentenormal = easyxf('font: name Calibri, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
                    style2 = easyxf('borders: left thin, right thin, top thin, bottom thin; alignment: horiz centre')
                    wb = xlwt.Workbook(encoding='utf-8')
                    ws = wb.add_sheet('postulantes')
                    ws.write_merge(0, 0, 0, 8, 'UNIVERSIDAD ESTATAL ESTATAL DE MILAGRO', title)
                    ws.write_merge(1, 1, 0, 8, 'LISTADO DE POSTULANTES BANCO HABILITANTES', title2)
                    row_num = 2
                    columns = [
                        ('Convocatoria', 10000),
                        ('Categoria', 10000),
                        ('Partida', 30000),
                        ('Carrera', 30000),
                        ('Asignaturas', 30000),
                        ('Dedicación', 10000),
                        ('RMU', 10000),
                        ('Nivel', 10000),
                        ('Modalidad', 10000),
                        ('Jornada', 10000),
                        ('Campo Amplio', 30000),
                        ('Campo Especifico', 30000),
                        ('Campo Detallado', 30000),
                        ('Cod. Unico', 10000),
                        ('Apellidos', 10000),
                        ('Nombres', 10000),
                        ('Titulo', 25000),
                        ('Archivo', 10000),
                        ('Identificación', 10000),
                        ('Correo', 10000),
                        ('Telf.', 10000),
                        ('Telf. Conv.', 10000),
                        ('Estado Calificación', 10000),
                        ('¿Tiene Formación?', 10000),
                        ('¿Tiene Experiencia?', 10000),
                        ('¿Tiene Capacitación?', 10000),
                        ('¿Tiene Publicaciones?', 10000),
                        ('¿Tiene Certificación Idiomas?', 10000),
                        ('Nota Formación', 10000),
                        ('Obs. Formación', 20000),
                        ('Nota Exp. Docente', 10000),
                        ('Obs. Exp. Docente', 20000),
                        ('Nota Exp. Administrativo', 10000),
                        ('Obs. Exp. Administrativo', 20000),
                        ('Nota Capacitación', 10000),
                        ('Obs. Capacitación', 20000),
                        ('¿Aplico Desempate?', 10000),
                        ('Nota Desempate', 10000),
                        ('Nota Final', 10000),
                        ('Obs. Final', 20000),
                        ('Estado', 10000),
                        ('¿Calificada?', 10000),
                        ('¿Apelo?', 10000),
                        ('Estado Apelación', 10000),
                        ('Obs. Apelación', 20000),
                        ('Usuario Revisión Meritos', 10000),
                        ('Fecha Revisión Meritos', 10000),
                        ('Usuario Revisión Apelación', 10000),
                        ('Fecha Revisión Apelación', 10000),
                        ('Usuario Revisión Desempate', 10000),
                        ('Fecha Revisión Desempate', 10000),
                    ]
                    font_style = xlwt.XFStyle()
                    font_style.font.bold = True
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], fuentecabecera)
                        ws.col(col_num).width = columns[col_num][1]
                    row_num += 1
                    convocatoria = Convocatoria.objects.get(status=True,pk=int(request.GET['id']))
                    for partidas in Partida.objects.filter(status=True, convocatoria=convocatoria).order_by('convocatoria__descripcion'):
                        listado = partidas.personaaplicarpartida_set.filter(status=True).order_by('-nota_final_meritos')[3:]
                        for det in listado:
                            if det.persona.mis_titulaciones().count() > 0:
                                row_count = row_num + det.persona.mis_titulaciones().count() - 1
                                titulos = det.persona.mis_titulaciones()
                                ws.write_merge(row_num, row_count, 0, 0, det.partida.convocatoria.descripcion, style2)
                                ws.write_merge(row_num, row_count, 1, 1, det.partida.convocatoria.tipocontrato.nombre if det.partida.convocatoria.tipocontrato else '', style2)
                                ws.write_merge(row_num, row_count, 2, 2, str(det.partida), style2)
                                ws.write_merge(row_num, row_count, 3, 3, det.partida.carrera.__str__() if det.partida.carrera else '', style2)
                                asignaturas_ = ''
                                for asig in det.partida.partidas_asignaturas():
                                    asignaturas_ += '{}, '.format(asig.__str__())
                                ws.write_merge(row_num, row_count, 4, 4, asignaturas_, style2)
                                ws.write_merge(row_num, row_count, 5, 5, det.partida.get_dedicacion_display(), style2)
                                ws.write_merge(row_num, row_count, 6, 6, det.partida.rmu, style2)
                                ws.write_merge(row_num, row_count, 7, 7, det.partida.get_nivel_display(), style2)
                                ws.write_merge(row_num, row_count, 8, 8, det.partida.get_modalidad_display(), style2)
                                ws.write_merge(row_num, row_count, 9, 9, det.partida.get_jornada_display(), style2)
                                campo_amplio = ''
                                for ca in det.partida.campoamplio.all():
                                    campo_amplio += '{}, '.format(ca.__str__())
                                ws.write_merge(row_num, row_count, 10, 10, campo_amplio, style2)
                                campo_especifico = ''
                                for ca in det.partida.campoespecifico.all():
                                    campo_especifico += '{}, '.format(ca.__str__())
                                ws.write_merge(row_num, row_count, 11, 11, campo_especifico, style2)
                                campo_detallado = ''
                                for ca in det.partida.campodetallado.all():
                                    campo_detallado += '{}, '.format(ca.__str__())
                                ws.write_merge(row_num, row_count, 12, 12, campo_detallado, style2)
                                ws.write_merge(row_num, row_count, 13, 13, det.pk, style2)
                                ws.write_merge(row_num, row_count, 14, 14, "{} {}".format(det.persona.apellido1, det.persona.apellido2), style2)
                                ws.write_merge(row_num, row_count, 15, 15, "{}".format(det.persona.nombres), style2)

                                ws.write_merge(row_num, row_count, 18, 18, det.persona.cedula, style2)
                                ws.write_merge(row_num, row_count, 19, 19, det.persona.email, style2)
                                ws.write_merge(row_num, row_count, 20, 20, det.persona.telefono, style2)
                                ws.write_merge(row_num, row_count, 21, 21, det.persona.telefono_conv, style2)
                                ws.write_merge(row_num, row_count, 22, 22, det.get_estado_display(), style2)
                                ws.write_merge(row_num, row_count, 23, 23, 'SI' if det.tiene_formacionacademica() else 'NO', style2)
                                ws.write_merge(row_num, row_count, 24, 24, 'SI' if det.tiene_experienciapartida() else 'NO', style2)
                                ws.write_merge(row_num, row_count, 25, 25, 'SI' if det.tiene_capacitaciones() else 'NO', style2)
                                ws.write_merge(row_num, row_count, 26, 26, 'SI' if det.tiene_publicaciones() else 'NO', style2)
                                ws.write_merge(row_num, row_count, 27, 27, 'SI' if det.tiene_idiomas() else 'NO', style2)
                                ws.write_merge(row_num, row_count, 28, 28, det.pgradoacademico, style2)
                                ws.write_merge(row_num, row_count, 29, 29, det.obsgradoacademico, style2)
                                ws.write_merge(row_num, row_count, 30, 30, det.pexpdocente, style2)
                                ws.write_merge(row_num, row_count, 31, 31, det.obsexperienciadoc, style2)
                                ws.write_merge(row_num, row_count, 32, 32, det.pexpadministrativa, style2)
                                ws.write_merge(row_num, row_count, 33, 33, det.obsexperienciaadmin, style2)
                                ws.write_merge(row_num, row_count, 34, 34, det.pcapacitacion, style2)
                                ws.write_merge(row_num, row_count, 35, 35, det.obscapacitacion, style2)
                                ws.write_merge(row_num, row_count, 36, 36, 'SI' if det.aplico_desempate else 'NO', style2)
                                ws.write_merge(row_num, row_count, 37, 37, det.nota_desempate, style2)
                                ws.write_merge(row_num, row_count, 38, 38, det.nota_final_meritos, style2)
                                ws.write_merge(row_num, row_count, 39, 39, det.obsgeneral, style2)
                                ws.write_merge(row_num, row_count, 40, 40, det.get_estado_display(), style2)
                                ws.write_merge(row_num, row_count, 41, 41, 'SI' if det.calificada else 'NO', style2)
                                ws.write_merge(row_num, row_count, 42, 42, 'SI' if det.solapelacion else 'NO', style2)
                                ws.write_merge(row_num, row_count, 43, 43, det.traer_apelacion().get_estado_display() if det.traer_apelacion() else '', style2)
                                obs_revisor, revisado_por, fecha_revision = '', '', ''
                                if det.traer_apelacion():
                                    if det.traer_apelacion().estado != 0:
                                        obs_revisor = det.traer_apelacion().observacion_revisor
                                        revisado_por = det.traer_apelacion().revisado_por.username if det.traer_apelacion().revisado_por else ''
                                        fecha_revision = str(det.traer_apelacion().fecha_revision)
                                ws.write_merge(row_num, row_count, 44, 44, obs_revisor, style2)
                                ws.write_merge(row_num, row_count, 45, 45, revisado_por, style2)
                                ws.write_merge(row_num, row_count, 46, 46, fecha_revision, style2)
                                ws.write_merge(row_num, row_count, 47, 47, det.revisado_por.username if det.revisado_por else '', style2)
                                ws.write_merge(row_num, row_count, 48, 48, str(det.fecha_revision) if det.fecha_revision else '', style2)
                                ws.write_merge(row_num, row_count, 49, 49, det.desempate_revisado_por.username if det.desempate_revisado_por else '', style2)
                                ws.write_merge(row_num, row_count, 50, 50, str(det.desempate_fecha_revision) if det.desempate_fecha_revision else '', style2)
                                for dato in titulos:
                                    ws.write(row_num, 16, "{}".format(dato.titulo), style2)
                                    ws.write(row_num, 17, "https://sga.unemi.edu.ec/{}".format(dato.archivo.url) if dato.archivo else '', style2)
                                    row_num += 1
                            else:
                                ws.write(row_num, 0, det.partida.convocatoria.descripcion, style2)
                                ws.write(row_num, 1, det.partida.convocatoria.tipocontrato.nombre if det.partida.convocatoria.tipocontrato else '', style2)
                                ws.write(row_num, 2, det.partida.titulo, style2)
                                ws.write(row_num, 3, det.partida.carrera.__str__() if det.partida.carrera else '', style2)
                                asignaturas_ = ''
                                for asig in det.partida.partidas_asignaturas():
                                    asignaturas_ += '{}, '.format(asig.__str__())
                                ws.write(row_num, 4, asignaturas_, style2)
                                ws.write(row_num, 5, det.partida.get_dedicacion_display(), style2)
                                ws.write(row_num, 6, det.partida.rmu, style2)
                                ws.write(row_num, 7, det.partida.get_nivel_display(), style2)
                                ws.write(row_num, 8, det.partida.get_modalidad_display(), style2)
                                ws.write(row_num, 9, det.partida.get_jornada_display(), style2)
                                campo_amplio = ''
                                for ca in det.partida.campoamplio.all():
                                    campo_amplio += '{}, '.format(ca.__str__())
                                ws.write(row_num, 10, campo_amplio, style2)
                                campo_especifico = ''
                                for ca in det.partida.campoespecifico.all():
                                    campo_especifico += '{}, '.format(ca.__str__())
                                ws.write(row_num, 11, campo_especifico, style2)
                                campo_detallado = ''
                                for ca in det.partida.campodetallado.all():
                                    campo_detallado += '{}, '.format(ca.__str__())
                                ws.write(row_num, 12, campo_detallado, style2)
                                ws.write(row_num, 13, det.pk, style2)
                                ws.write(row_num, 14, "{} {}".format(det.persona.apellido1, det.persona.apellido2), style2)
                                ws.write(row_num, 15, "{}".format(det.persona.nombres), style2)
                                ws.write(row_num, 16, "", style2)
                                ws.write(row_num, 17, "", style2)
                                ws.write(row_num, 18, det.persona.cedula, style2)
                                ws.write(row_num, 19, det.persona.email, style2)
                                ws.write(row_num, 20, det.persona.telefono, style2)
                                ws.write(row_num, 21, det.persona.telefono_conv, style2)
                                ws.write(row_num, 22, det.get_estado_display(), style2)
                                ws.write(row_num, 23, 'SI' if det.tiene_formacionacademica() else 'NO', style2)
                                ws.write(row_num, 24, 'SI' if det.tiene_experienciapartida() else 'NO', style2)
                                ws.write(row_num, 25, 'SI' if det.tiene_capacitaciones() else 'NO', style2)
                                ws.write(row_num, 26, 'SI' if det.tiene_publicaciones() else 'NO', style2)
                                ws.write(row_num, 27, 'SI' if det.tiene_idiomas() else 'NO', style2)
                                ws.write(row_num, 28, det.pgradoacademico, style2)
                                ws.write(row_num, 29, det.obsgradoacademico, style2)
                                ws.write(row_num, 30, det.pexpdocente, style2)
                                ws.write(row_num, 31, det.obsexperienciadoc, style2)
                                ws.write(row_num, 32, det.pexpadministrativa, style2)
                                ws.write(row_num, 33, det.obsexperienciaadmin, style2)
                                ws.write(row_num, 34, det.pcapacitacion, style2)
                                ws.write(row_num, 35, det.obscapacitacion, style2)
                                ws.write(row_num, 36, 'SI' if det.aplico_desempate else 'NO', style2)
                                ws.write(row_num, 37, det.nota_desempate, style2)
                                ws.write(row_num, 38, det.nota_final_meritos, style2)
                                ws.write(row_num, 39, det.obsgeneral, style2)
                                ws.write(row_num, 40, det.get_estado_display(), style2)
                                ws.write(row_num, 41, 'SI' if det.calificada else 'NO', style2)
                                ws.write(row_num, 42, 'SI' if det.solapelacion else 'NO', style2)
                                ws.write(row_num, 43, det.traer_apelacion().get_estado_display() if det.traer_apelacion() else '', style2)
                                obs_revisor, revisado_por, fecha_revision = '', '', ''
                                if det.traer_apelacion():
                                    if det.traer_apelacion().estado != 0:
                                        obs_revisor = det.traer_apelacion().observacion_revisor
                                        revisado_por = det.traer_apelacion().revisado_por.username if det.traer_apelacion().revisado_por else ''
                                        fecha_revision = str(det.traer_apelacion().fecha_revision)
                                ws.write(row_num, 44, obs_revisor, style2)
                                ws.write(row_num, 45, revisado_por, style2)
                                ws.write(row_num, 46, fecha_revision, style2)
                                ws.write(row_num, 47, det.revisado_por.username if det.revisado_por else '', style2)
                                ws.write(row_num, 48, str(det.fecha_revision) if det.fecha_revision else '', style2)
                                ws.write(row_num, 49, det.desempate_revisado_por.username if det.desempate_revisado_por else '', style2)
                                ws.write(row_num, 50, str(det.desempate_fecha_revision) if det.desempate_fecha_revision else '', style2)
                                row_num += 1
                    wb.save(response)
                    return response
                except Exception as ex:
                    msg_ex = 'Error on line {} - {}'.format(sys.exc_info()[-1].tb_lineno, str(ex))
                    return JsonResponse({"result": False, 'data': str(msg_ex)})

            if action == 'vercalificar':
                try:
                    data['id'] = id = int(encrypt(request.GET['id']))
                    data['postulante'] = postulante = PersonaAplicarPartida.objects.get(pk=id)
                    data['partida'] = partida = Partida.objects.get(pk=postulante.partida.pk)
                    data['resp_campos'] = validar_campos(request, persona, partida)
                    data['persona'] = postulante.persona
                    data['posidiomas'] = posidiomas = PersonaIdiomaPartida.objects.filter(status=True, personapartida=postulante).order_by('id')
                    data['postitulacion'] = postitulacion = PersonaFormacionAcademicoPartida.objects.filter(status=True, personapartida=postulante).order_by('id')
                    data['posexperiencia'] = posexperiencia = PersonaExperienciaPartida.objects.filter(status=True, personapartida=postulante).order_by('id')
                    data['poscapacitacion'] = poscapacitacion = PersonaCapacitacionesPartida.objects.filter(status=True, personapartida=postulante).order_by('id')
                    data['pospublicacion'] = pospublicacion = PersonaPublicacionesPartida.objects.filter(status=True, personapartida=postulante).order_by('id')
                    template = get_template("postulate/mispostulaciones/vercalificacion.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            if action == 'reportdesiertos':
                try:
                    sin_pos = Partida.objects.values_list('id', flat=True).filter(status=True,vigente=True,personaaplicarpartida__isnull=True)
                    pos_menor = PersonaAplicarPartida.objects.values_list('partida_id', flat=True).filter(status=True, nota_final__lt=70).distinct('partida_id')
                    re = sin_pos.union(pos_menor)
                    re = list(re)
                    listado = Partida.objects.filter(status=True, vigente=True, id__in=re).distinct('id')

                    __autor__ = 'unemi'
                    output = io.BytesIO()
                    workbook = xlsxwriter.Workbook(output)
                    ws = workbook.add_worksheet('partidas')

                    ws.set_row(0,20)
                    title = workbook.add_format({'font_name':'Calibri','align':'center','bold':True,'font_size':18,'valign':'vcenter'})
                    title2 = workbook.add_format({'font_name':'Calibri','align':'center','bold':True,'font_size':14,'valign':'vcenter'})
                    fuentecabecera = workbook.add_format({'font_name':'Calibri','border':1,'align':'center','font_size':11,'valign':'vcenter','bg_color':'#E4E5DF'})
                    style2 = workbook.add_format({'text_wrap':True,'font_name': 'Calibri', 'border': 1, 'align': 'center', 'font_size': 11, 'valign': 'vcenter'})
                    ws.merge_range(0, 0, 0, 8, 'UNIVERSIDAD ESTATAL ESTATAL DE MILAGRO', title)
                    ws.merge_range(1, 1, 0, 8, '', title2)
                    row_num = 2
                    columns = [
                        ('Cod. Partida', 15),
                        ('Título', 120),
                        ('Descripción', 120),
                        ('Carrera', 80),
                        ('Campo Amplio', 120),
                        ('Campo Especifico', 100),
                        ('Campo Detallado', 120),
                        ('Títulos Relacionados', 120),
                        ('Nivel', 35),
                        ('Modalidad', 30),
                        ('Dedicación', 40),
                        ('Jornada', 23),
                        ('RMU', 10),
                        ('Vigente', 10),
                        ('Total Postulantes', 10), ]
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], fuentecabecera)
                        ws.set_column(col_num,col_num,columns[col_num][1])
                    row_num += 1
                    for det in listado:
                        if det.participantes_mejores_puntuados():
                            pass
                        else:
                            ws.write(row_num, 0, det.codpartida, style2)
                            ws.write(row_num, 1, str(det), style2)
                            # ws.write(row_num, 2, det.descripcion, style2)
                            ws.write(row_num, 2, det.carrera.__str__() if det.carrera else '', style2)
                            campo_amplio = ''
                            for ca in det.campoamplio.all():
                                campo_amplio += '{}, '.format(ca.__str__())
                            ws.write(row_num, 3, campo_amplio, style2)
                            campo_especifico = ''
                            for ca in det.campoespecifico.all():
                                campo_especifico += '{}, '.format(ca.__str__())
                            ws.write(row_num, 4, campo_especifico, style2)
                            campo_detallado = ''
                            for ca in det.campodetallado.all():
                                campo_detallado += '{}, '.format(ca.__str__())
                            ws.write(row_num, 5, campo_detallado, style2)
                            titulos_relacionados = ''
                            for ca in det.titulos.all():
                                titulos_relacionados += '{}, '.format(ca.__str__())
                            ws.write(row_num, 6, titulos_relacionados, style2)
                            ws.write(row_num, 7, det.get_nivel_display(), style2)
                            ws.write(row_num, 8, det.get_modalidad_display(), style2)
                            ws.write(row_num, 9, det.get_dedicacion_display(), style2)
                            ws.write(row_num, 10, det.get_jornada_display(), style2)
                            ws.write(row_num, 11, det.rmu, style2)
                            ws.write(row_num, 12, 'SI' if det.vigente else 'NO', style2)
                            ws.write(row_num, 13, det.total_postulantes(), style2)
                            row_num += 1
                    workbook.close()
                    output.seek(0)
                    filename = 'reporte_desiertos' + random.randint(1, 10000).__str__() + '.xlsx'
                    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    response['Content-Disposition'] = 'attachment; filename=%s' % filename
                    return response
                except Exception as ex:
                    pass

            if action == 'reporttribunal':
                try:
                    partidas = Partida.objects.filter(status=True,vigente=True)
                    # tribunal = partidas.partidatribunal_set.filter(tipo =1)
                    etapa = int(request.GET['etapa'])

                    __autor__ = 'unemi'
                    output = io.BytesIO()
                    workbook = xlsxwriter.Workbook(output)
                    ws = workbook.add_worksheet('partidas_tribunal')
                    ws.set_column(0,1,150)
                    ws.set_row(0,20)

                    title = workbook.add_format({'font_name': 'Calibri', 'align': 'center', 'bold': True, 'font_size': 18, 'valign': 'vcenter'})
                    title2 = workbook.add_format({'font_name': 'Calibri', 'align': 'center', 'bold': True, 'font_size': 14, 'valign': 'vcenter'})
                    fuentecabecera = workbook.add_format({'font_name': 'Calibri', 'border': 1, 'align': 'center', 'font_size': 11, 'valign': 'vcenter', 'bg_color': '#E4E5DF'})
                    style2 = workbook.add_format({'text_wrap': True, 'font_name': 'Calibri', 'border': 1, 'align': 'center', 'font_size': 11, 'valign': 'vcenter'})
                    ws.merge_range(0, 0, 0, 8, 'UNIVERSIDAD ESTATAL ESTATAL DE MILAGRO', title)
                    ws.merge_range(1, 1, 0, 8, '', title2)
                    row_num =3
                    columns = [
                        ('Convocatoria', 80),
                        ('Cod. Partida', 15),
                        ('Denominación de puesto', 90),
                    ]
                    for cargo in CARGOS_TRIBUNAL:
                        columns.append((cargo[1],70))
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], fuentecabecera)
                        ws.set_column(col_num,col_num,columns[col_num][1])
                    ws.merge_range(row_num-1,3,row_num-1,3+len(CARGOS_TRIBUNAL)-1,'CARGOS TRIBUNAL',fuentecabecera)
                    row_num+=1
                    for par in partidas:
                        ws.write(row_num,0,par.convocatoria.__str__(),style2)
                        ws.write(row_num,1,par.codpartida,style2)
                        ws.write(row_num,2,(par.denominacionpuesto if par.denominacionpuesto else ''),style2)
                        aux = 2
                        for l in par.cargos_tribunal(etapa):
                            ws.write(row_num,aux+int(CARGOS_TRIBUNAL[l.cargos-1][0]),l.persona.__str__() if l.persona else ' ',style2)
                        row_num+=1

                    workbook.close()
                    output.seek(0)
                    filename = 'reporte_tribunal_partidas_' + random.randint(1, 10000).__str__() + '.xlsx'
                    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    response['Content-Disposition'] = 'attachment; filename=%s' % filename
                    return response
                except Exception as ex:
                    pass

            if action == 'reportdesiertos':
                try:
                    sin_pos = Partida.objects.values_list('id', flat=True).filter(status=True,vigente=True,personaaplicarpartida__isnull=True)
                    pos_menor = PersonaAplicarPartida.objects.values_list('partida_id', flat=True).filter(status=True, nota_final__lt=70).distinct('partida_id')
                    re = sin_pos.union(pos_menor)
                    re = list(re)
                    listado = Partida.objects.filter(status=True, vigente=True, id__in=re).distinct('id')

                    __autor__ = 'unemi'
                    output = io.BytesIO()
                    workbook = xlsxwriter.Workbook(output)
                    ws = workbook.add_worksheet('partidas')

                    ws.set_row(0,20)
                    title = workbook.add_format({'font_name':'Calibri','align':'center','bold':True,'font_size':18,'valign':'vcenter'})
                    title2 = workbook.add_format({'font_name':'Calibri','align':'center','bold':True,'font_size':14,'valign':'vcenter'})
                    fuentecabecera = workbook.add_format({'font_name':'Calibri','border':1,'align':'center','font_size':11,'valign':'vcenter','bg_color':'#E4E5DF'})
                    style2 = workbook.add_format({'text_wrap':True,'font_name': 'Calibri', 'border': 1, 'align': 'center', 'font_size': 11, 'valign': 'vcenter'})
                    ws.merge_range(0, 0, 0, 8, 'UNIVERSIDAD ESTATAL ESTATAL DE MILAGRO', title)
                    ws.merge_range(1, 1, 0, 8, '', title2)
                    row_num = 2
                    columns = [
                        ('Cod. Partida', 15),
                        ('Título', 120),
                        ('Descripción', 120),
                        ('Carrera', 80),
                        ('Campo Amplio', 120),
                        ('Campo Especifico', 100),
                        ('Campo Detallado', 120),
                        ('Títulos Relacionados', 120),
                        ('Nivel', 35),
                        ('Modalidad', 30),
                        ('Dedicación', 40),
                        ('Jornada', 23),
                        ('RMU', 10),
                        ('Vigente', 10),
                        ('Total Postulantes', 10), ]
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], fuentecabecera)
                        ws.set_column(col_num,col_num,columns[col_num][1])
                    row_num += 1
                    for det in listado:
                        if det.participantes_mejores_puntuados():
                            pass
                        else:
                            ws.write(row_num, 0, det.codpartida, style2)
                            ws.write(row_num, 1, det.titulo, style2)
                            ws.write(row_num, 2, det.descripcion, style2)
                            ws.write(row_num, 3, det.carrera.__str__() if det.carrera else '', style2)
                            campo_amplio = ''
                            for ca in det.campoamplio.all():
                                campo_amplio += '{}, '.format(ca.__str__())
                            ws.write(row_num, 4, campo_amplio, style2)
                            campo_especifico = ''
                            for ca in det.campoespecifico.all():
                                campo_especifico += '{}, '.format(ca.__str__())
                            ws.write(row_num, 5, campo_especifico, style2)
                            campo_detallado = ''
                            for ca in det.campodetallado.all():
                                campo_detallado += '{}, '.format(ca.__str__())
                            ws.write(row_num, 6, campo_detallado, style2)
                            titulos_relacionados = ''
                            for ca in det.titulos.all():
                                titulos_relacionados += '{}, '.format(ca.__str__())
                            ws.write(row_num, 7, titulos_relacionados, style2)
                            ws.write(row_num, 8, det.get_nivel_display(), style2)
                            ws.write(row_num, 9, det.get_modalidad_display(), style2)
                            ws.write(row_num, 10, det.get_dedicacion_display(), style2)
                            ws.write(row_num, 11, det.get_jornada_display(), style2)
                            ws.write(row_num, 12, det.rmu, style2)
                            ws.write(row_num, 13, 'SI' if det.vigente else 'NO', style2)
                            ws.write(row_num, 14, det.total_postulantes(), style2)
                            row_num += 1
                    workbook.close()
                    output.seek(0)
                    filename = 'reporte_desiertos' + random.randint(1, 10000).__str__() + '.xlsx'
                    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    response['Content-Disposition'] = 'attachment; filename=%s' % filename
                    return response
                except Exception as ex:
                    pass

            if action == 'verdetallepostulante':
                try:
                    data['id'] = id = int(encrypt(request.GET['id']))
                    data['partida'] = partida = Partida.objects.get(pk=id, status=True)
                    data['postulante'] = postulante = PersonaAplicarPartida.objects.get(partida = partida, esganador=True,finsegundaetapa=True)
                    data['resp_campos'] = validar_campos(request, persona, partida)
                    data['persona'] = postulante.persona
                    data['posidiomas'] = PersonaIdiomaPartida.objects.filter(status=True, personapartida=postulante).order_by('id')
                    data['postitulacion'] = PersonaFormacionAcademicoPartida.objects.filter(status=True, personapartida=postulante).order_by('id')
                    data['posexperiencia'] = PersonaExperienciaPartida.objects.filter(status=True, personapartida=postulante).order_by('id')
                    data['poscapacitacion'] = PersonaCapacitacionesPartida.objects.filter(status=True, personapartida=postulante).order_by('id')
                    data['pospublicacion'] = PersonaPublicacionesPartida.objects.filter(status=True, personapartida=postulante).order_by('id')
                    template = get_template("postulate/adm_revisionpostulacion/verdetallepostulante.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            if action == 'duplicarconvocatoria':
                try:
                    if 'id' in request.GET:
                        id = int(encrypt(request.GET['id']))
                    if Convocatoria.objects.values('id').filter(pk=id,status=True):
                        convocatoria = Convocatoria.objects.get(pk =id, status=True)
                        conv = Convocatoria(descripcion='Copia - {}'.format(convocatoria.descripcion),
                                                finicio=convocatoria.finicio,
                                                ffin=convocatoria.ffin,
                                                tipocontrato=convocatoria.tipocontrato,
                                                denominacionpuesto=convocatoria.denominacionpuesto,
                                                modeloevaluativo = convocatoria.modeloevaluativo,
                                                vigente=convocatoria.vigente)
                        conv.valtercernivel = convocatoria.valtercernivel
                        conv.valposgrado = convocatoria.valposgrado
                        conv.valdoctorado = convocatoria.valdoctorado
                        conv.valcapacitacionmin = convocatoria.valcapacitacionmin
                        conv.valcapacitacionmax = convocatoria.valcapacitacionmax
                        conv.valexpdocentemin = convocatoria.valexpdocentemin
                        conv.valexpdocentemax = convocatoria.valexpdocentemax
                        conv.valexpadminmin = convocatoria.valexpadminmin
                        conv.valexpadminmax = convocatoria.valexpadminmax
                        conv.save(request)
                        for condi in convocatoria.convocatoriaterminoscondiciones_set.filter(status=True):
                            termino = ConvocatoriaTerminosCondiciones(
                                convocatoria = conv,
                                descripcion = condi.descripcion
                            )
                            termino.save(request)
                        log('Ha duplicado la convocatoria {}, con terminos y condiciones'.format(conv.__str__()),request,'change')
                        return JsonResponse({"result": True, 'msg': 'Convocatoria Duplicada!'})
                    else:
                        return JsonResponse({"result":False,'msg':'No existe la convocatoria'})
                except Exception as ex:
                    pass

            if action == 'duplicarpartidaplanificacion':
                try:
                    if 'id' in request.GET:
                        id = int(encrypt(request.GET['id']))
                    if 'idp' in request.GET:
                        idp = int(encrypt(request.GET['idp']))
                    if Partida.objects.values('id').filter(pk=id, status=True, periodoplanificacion_id=idp):
                        partida = Partida.objects.get(pk=id, status=True, periodoplanificacion_id=idp)
                        partidaasignaturas = PartidaAsignaturas.objects.filter(partida=partida, status=True).values_list('asignatura_id')
                        detallecapacitacion = DetalleCapacitacionPlanificacion.objects.filter(partida=partida, status=True)
                        part = Partida(
                                        periodoplanificacion_id = idp,
                                        codpartida=partida.codpartida,
                                        titulo = partida.titulo,
                                        descripcion='Copia - {}'.format(partida.descripcion),
                                        denominacionpuesto = partida.denominacionpuesto,
                                        carrera=partida.carrera,
                                        nivel=partida.nivel,
                                        modalidad=partida.modalidad,
                                        dedicacion=partida.dedicacion,
                                        jornada=partida.jornada,
                                        rmu=partida.rmu,
                                        estado = partida.estado,
                                        temadisertacion = partida.temadisertacion,
                                        observacion = partida.observacion)
                        part.save()
                        for ca in partida.campoamplio.all():
                            part.campoamplio.add(ca)
                        for caesp in partida.campoespecifico.all():
                            part.campoespecifico.add(caesp)
                        for cadet in partida.campodetallado.all():
                            part.campodetallado.add(cadet)
                        for tit in partida.titulos.all():
                            part.titulos.add(tit)
                        part.save(request)
                        for asignatura in partidaasignaturas:
                            partasignatura = PartidaAsignaturas(partida=part, asignatura_id=asignatura[0])
                            partasignatura.save(request)
                        for detalle in detallecapacitacion:
                            det = DetalleCapacitacionPlanificacion(
                                    partida = part,
                                    tipocompetencia_id=detalle.tipocompetencia_id,
                                    tiempocapacitacion=detalle.tiempocapacitacion,
                                    canttiempocapacitacion=detalle.canttiempocapacitacion,
                                    cespecifica_id=detalle.cespecifica_id if detalle.cespecifica_id != 0 else None,
                                    descripcioncapacitacion=detalle.descripcioncapacitacion
                            )
                            det.save(request)
                        log('Ha duplicado la partida {}, con asignaturas'.format(part.__str__()),
                            request, 'change')
                        return JsonResponse({"result": True, 'msg': 'Partida Duplicada!'})
                    else:
                        return JsonResponse({"result": False, 'msg': 'No existe la partida'})
                except Exception as ex:
                    pass

            elif action == 'addperiodo':
                try:
                    data['title'] = u'Adicionar Periodo'
                    form = PeriodoAcademicoConvocatoriaForm()
                    data['form'] = form
                    template = get_template("postulate/adm_periodoplanificacion/modal/formperiodo.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'validarpartida':
                try:
                    data['title'] = u'Validar partida'
                    form = ValidarPartidaPlanificacionForm()
                    data['id'] = encrypt(request.GET['id'])
                    data['form'] = form
                    template = get_template("postulate/adm_periodoplanificacion/modal/formaprobar.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'requisitos':
                try:
                    data['title'] = u'Requisitos para contratos'
                    search, url_vars, filtro = request.GET.get('s', ''), '',Q(status=True)
                    requisito = RequisitoDocumentoContrato.objects.filter(status=True)
                    if search:
                        data['search'] = search
                        url_vars += "&s={}".format(search)
                        filtro = filtro & (Q(descripcion__icontains=search)|Q(nombre__icontains=search))
                        requisito = requisito.filter(filtro).order_by('nombre','descripcion')
                    paging = MiPaginador(requisito, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data["url_vars"] = url_vars
                    data['listado'] = page.object_list
                    data['list_count'] = len(requisito)
                    return render(request, "postulate/adm_postulate/viewrequisito.html", data)
                except Exception as ex:
                    pass

            elif action == 'addrequisito':
                try:
                    data['title'] = u'Adiccionar Requisito'
                    form = RequisitoDocumentoContratoForm()
                    data['form'] = form
                    template = get_template("postulate/adm_postulate/modal/formrequisito.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass
            elif action == 'editrequisito':
                try:
                    data['title'] = u'Editar Requisito'
                    data['requisito'] = filtro = RequisitoDocumentoContrato.objects.get(id = int(encrypt(request.GET['id'])))
                    form = RequisitoDocumentoContratoForm(initial = {
                        'nombre':filtro.nombre,
                        'descripcion':filtro.descripcion,
                        'anio':filtro.anio,
                        'tipo':filtro.tipo,
                        'obligatorio':filtro.obligatorio,
                        'activo':filtro.activo,
                        'archivo':filtro.archivo
                    })
                    data['form'] = form
                    template = get_template("postulate/adm_postulate/modal/formrequisito.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass
            elif action == "ganadores":
                try:
                    data['title'] = 'Ganadores'
                    data['personaaplica'] = personaaplica = PersonaAplicarPartida.objects.filter(status=True,esganador=True)
                    search, url_vars, filtro = request.GET.get('s', ''), '', Q(status=True)
                    if search:
                        data['search'] = search
                        s = search.split(' ')
                        url_vars += "&s={}".format(search)
                        if len(s)>1:
                            filtro = filtro & (Q(persona__apellido1__icontains=s[0]) | Q(persona__apellido2__icontains=s[1]))
                        else:
                            filtro = filtro & (Q(persona__cedula__icontains=s[0]))
                        personaaplica = personaaplica.filter(filtro).order_by('id')
                    paging = MiPaginador(personaaplica, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data["url_vars"] = url_vars
                    data['listado'] = page.object_list
                    return render(request, "postulate/adm_postulate/viewganadores.html", data)
                except Exception as ex:
                    pass

            elif action == 'viewticompeplanificacion':
                try:
                    data['title'] = 'Tipo Competencia Planificación'
                    data['tipocompeplani'] = tipocompeplani = TipoCompetenciaPlanificacion.objects.filter(status=True).order_by('id')
                    search, url_vars, filtro = request.GET.get('s', ''), '', Q(status=True)
                    if search:
                        data['search'] = search
                        url_vars += "&s={}".format(search)
                        filtro = filtro & (Q(nombre__icontains=search))
                        tipocompeplani = tipocompeplani.filter(filtro).order_by('id')
                    paging = MiPaginador(tipocompeplani, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    request.session['viewactivoperplanificacion'] = 1
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data["url_vars"] = url_vars
                    data['listado'] = page.object_list
                    return render(request, "postulate/adm_periodoplanificacion/viewticompeplanificacion.html", data)
                except Exception as ex:
                    pass

            elif action == 'viewticompeespeplanificacion':
                try:
                    data['title'] = 'Tipo Competencia Específica Planificación'
                    data['tipocompeplani'] = tipocompeplani = TipoCompetenciaEspecificaPlanificacion.objects.filter(status=True).order_by('id')
                    search, url_vars, filtro = request.GET.get('s', ''), '', Q(status=True)
                    if search:
                        data['search'] = search
                        url_vars += "&s={}".format(search)
                        filtro = filtro & (Q(nombre__icontains=search))
                        tipocompeplani = tipocompeplani.filter(filtro).order_by('id')
                    paging = MiPaginador(tipocompeplani, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['viewactivoperplanificacion'] = 2
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data["url_vars"] = url_vars
                    data['listado'] = page.object_list
                    return render(request, "postulate/adm_periodoplanificacion/viewticompeespeplanificacion.html", data)
                except Exception as ex:
                    pass

            if action == 'reporte_postulantes_partidas':
                try:
                    periodo_id = request.GET.get('id') if request.GET.get('id') else 0
                    participantes = PersonaAplicarPartida.objects.filter(status=True, partida__periodoplanificacion_id=periodo_id).order_by('partida__convocatoria__descripcion')
                    __author__ = 'Unemi'
                    ahora = datetime.now()
                    time_codigo = ahora.strftime('%Y%m%d_%H%M%S')
                    name_file = f'reporte_excel_postulantes_puntuados_{time_codigo}.xlsx'
                    output = io.BytesIO()
                    workbook = xlsxwriter.Workbook(output)
                    ws = workbook.add_worksheet("Postulantes")

                    fuenteencabezado = workbook.add_format({
                        'align': 'center',
                        'bg_color': '#1C3247',
                        'font_color': 'white',
                        'border': 1,
                        'font_size': 24,
                        'bold': 1
                    })

                    fuentecabecera = workbook.add_format({
                        'align': 'center',
                        'bg_color': 'silver',
                        'border': 1,
                        'bold': 1
                    })

                    formatoceldacenter = workbook.add_format({
                        'border': 1,
                        'valign': 'vcenter',
                        'align': 'center'})


                    formatoceldafecha = workbook.add_format({
                        'num_format': 'dd/mm/yyyy',
                        'border': 1,
                        'valign': 'vcenter',
                        'align': 'center'
                    })

                    columnas = [
                        ('Período Planificación', 60),
                        ('Partida', 60),
                        ('Tema disertación', 30),
                        ('Carrera', 90),
                        ('Asignaturas', 90),
                        ('Dedicación', 30),
                        ('RMU', 10),
                        ('Nivel', 20),
                        ('Modalidad', 20),
                        ('Jornada', 20),
                        ('Campo Amplio', 65),
                        ('Campo Especifico', 65),
                        ('Campo Detallado', 65),
                        ('Cod. Unico', 10),
                        ('Apellidos', 20),
                        ('Nombres', 20),
                        ('Titulo', 140),
                        ('Archivo', 80),
                        ('Identificación', 20),
                        ('Correo', 40),
                        ('Telf.', 15),
                        ('Telf. Conv.', 15),
                        ('Estado Calificación', 20),
                        ('¿Tiene Formación?', 30),
                        ('¿Tiene Experiencia?', 30),
                        ('¿Tiene Capacitación?', 30),
                        ('¿Tiene Publicaciones?', 30),
                        ('¿Tiene Certificación Idiomas?', 30),
                        ('Nota Formación', 16),
                        ('Obs. Formación', 50),
                        ('Nota Exp. Docente', 22),
                        ('Obs. Exp. Docente', 22),
                        ('Nota Exp. Administrativo', 35),
                        ('Obs. Exp. Administrativo', 35),
                        ('Nota Capacitación', 35),
                        ('Obs. Capacitación', 35),
                        ('¿Aplico Desempate?', 30),
                        ('Nota Desempate', 20),
                        ('Nota Final', 15),
                        ('Obs. Final', 40),
                        ('Estado', 20),
                        ('¿Calificada?', 20),
                        ('¿Apelo?', 12),
                        ('Estado Apelación', 20),
                        ('Obs. Apelación', 40),
                        ('Usuario Revisión Meritos', 40),
                        ('Fecha Revisión Meritos', 40),
                        ('Usuario Revisión Apelación', 40),
                        ('Fecha Revisión Apelación', 40),
                        ('Usuario Revisión Desempate', 40),
                        ('Fecha Revisión Desempate', 40),
                    ]

                    ws.merge_range(0, 0, 0, 50, 'UNIVERSIDAD ESTATAL ESTATAL DE MILAGRO', fuenteencabezado)
                    ws.merge_range(1, 0, 1, 50, 'LISTADO DE POSTULANTES', fuenteencabezado)

                    row_num, numcolum = 3, 0

                    for col_name in columnas:
                        ws.write(row_num, numcolum, col_name[0], fuentecabecera)
                        ws.set_column(row_num, numcolum, col_name[1])
                        numcolum += 1

                    row_num += 1

                    for participante in participantes:
                        titulos = participante.persona.mis_titulaciones()
                        if titulos.count() > 1:
                            row_count = row_num + titulos.count() - 1
                            ws.merge_range(row_num, 0, row_count, 0, participante.partida.periodoplanificacion.nombre, formatoceldacenter)
                            ws.merge_range(row_num, 1, row_count, 1, str(participante.partida), formatoceldacenter)
                            ws.merge_range(row_num, 2, row_count, 2, participante.partida.temadisertacion, formatoceldacenter)
                            ws.merge_range(row_num, 3, row_count, 3, participante.partida.carrera.__str__() if participante.partida.carrera else '', formatoceldacenter)
                            asignaturas = ", ".join([asig.__str__() for asig in participante.partida.partidas_asignaturas()])
                            ws.merge_range(row_num, 4, row_count, 4, asignaturas, formatoceldacenter)
                            ws.merge_range(row_num, 5, row_count, 5, participante.partida.get_dedicacion_display(), formatoceldacenter)
                            ws.merge_range(row_num, 6, row_count, 6, participante.partida.rmu, formatoceldacenter)
                            ws.merge_range(row_num, 7, row_count, 7, participante.partida.get_nivel_display(), formatoceldacenter)
                            ws.merge_range(row_num, 8, row_count, 8, participante.partida.get_modalidad_display(), formatoceldacenter)
                            ws.merge_range(row_num, 9, row_count, 9, participante.partida.get_jornada_display(), formatoceldacenter)
                            campo_amplio = ", ".join([camp.__str__() for camp in participante.partida.campoamplio.all()])
                            ws.merge_range(row_num, 10, row_count, 10, campo_amplio, formatoceldacenter)
                            campo_especifico = ", ".join([camp.__str__() for camp in participante.partida.campoespecifico.all()])
                            ws.merge_range(row_num, 11, row_count, 11, campo_especifico, formatoceldacenter)
                            campo_detallado = ", ".join([camp.__str__() for camp in participante.partida.campodetallado.all()])
                            ws.merge_range(row_num, 12, row_count, 12, campo_detallado, formatoceldacenter)
                            ws.merge_range(row_num, 13, row_count, 13, participante.id, formatoceldacenter)
                            ws.merge_range(row_num, 14, row_count, 14, f'{participante.persona.apellido1} {participante.persona.apellido2}', formatoceldacenter)
                            ws.merge_range(row_num, 15, row_count, 15, f'{participante.persona.nombres}', formatoceldacenter)
                            ws.merge_range(row_num, 18, row_count, 18, participante.persona.cedula, formatoceldacenter)
                            ws.merge_range(row_num, 19, row_count, 19, participante.persona.email, formatoceldacenter)
                            ws.merge_range(row_num, 20, row_count, 20, participante.persona.telefono, formatoceldacenter)
                            ws.merge_range(row_num, 21, row_count, 21, participante.persona.telefono_conv, formatoceldacenter)
                            ws.merge_range(row_num, 22, row_count, 22, participante.get_estado_display(), formatoceldacenter)
                            ws.merge_range(row_num, 23, row_count, 23, 'SI' if participante.tiene_formacionacademica() else 'NO', formatoceldacenter)
                            ws.merge_range(row_num, 24, row_count, 24, 'SI' if participante.tiene_experienciapartida() else 'NO', formatoceldacenter)
                            ws.merge_range(row_num, 25, row_count, 25, 'SI' if participante.tiene_capacitaciones() else 'NO', formatoceldacenter)
                            ws.merge_range(row_num, 26, row_count, 26, 'SI' if participante.tiene_publicaciones() else 'NO', formatoceldacenter)
                            ws.merge_range(row_num, 27, row_count, 27, 'SI' if participante.tiene_idiomas() else 'NO', formatoceldacenter)
                            ws.merge_range(row_num, 28, row_count, 28, participante.pgradoacademico, formatoceldacenter)
                            ws.merge_range(row_num, 29, row_count, 29, participante.obsgradoacademico, formatoceldacenter)
                            ws.merge_range(row_num, 30, row_count, 30, participante.pexpdocente, formatoceldacenter)
                            ws.merge_range(row_num, 31, row_count, 31, participante.obsexperienciadoc, formatoceldacenter)
                            ws.merge_range(row_num, 32, row_count, 32, participante.pexpadministrativa, formatoceldacenter)
                            ws.merge_range(row_num, 33, row_count, 33, participante.obsexperienciaadmin, formatoceldacenter)
                            ws.merge_range(row_num, 34, row_count, 34, participante.pcapacitacion, formatoceldacenter)
                            ws.merge_range(row_num, 35, row_count, 35, participante.obscapacitacion, formatoceldacenter)
                            ws.merge_range(row_num, 36, row_count, 36, 'SI' if participante.aplico_desempate else 'NO', formatoceldacenter)
                            ws.merge_range(row_num, 37, row_count, 37, participante.nota_desempate, formatoceldacenter)
                            ws.merge_range(row_num, 38, row_count, 38, participante.nota_final_meritos, formatoceldacenter)
                            ws.merge_range(row_num, 39, row_count, 39, participante.obsgeneral, formatoceldacenter)
                            ws.merge_range(row_num, 40, row_count, 40, participante.get_estado_display(), formatoceldacenter)
                            ws.merge_range(row_num, 41, row_count, 41, 'SI' if participante.calificada else 'NO', formatoceldacenter)
                            ws.merge_range(row_num, 42, row_count, 42, 'SI' if participante.solapelacion else 'NO', formatoceldacenter)
                            participante_apelacion = participante.traer_apelacion()
                            ws.merge_range(row_num, 43, row_count, 43, participante_apelacion.get_estado_display() if participante_apelacion else "", formatoceldacenter)
                            obs_revisor, revisado_por, fecha_revision = '', '', ''
                            if participante_apelacion:
                                if participante_apelacion.estado != 0:
                                    obs_revisor = participante_apelacion.observacion_revisor
                                    revisado_por = participante_apelacion.revisado_por.username if participante_apelacion.revisado_por else ''
                                    fecha_revision = participante_apelacion.fecha_revision

                            ws.merge_range(row_num, 44, row_count, 44, obs_revisor, formatoceldacenter)
                            ws.merge_range(row_num, 45, row_count, 45, revisado_por, formatoceldacenter)
                            ws.merge_range(row_num, 46, row_count, 46, fecha_revision, formatoceldafecha)
                            ws.merge_range(row_num, 47, row_count, 47, participante.revisado_por.username if participante.revisado_por else '', formatoceldacenter)
                            ws.merge_range(row_num, 48, row_count, 48, participante.fecha_revision if participante.fecha_revision else '', formatoceldafecha)
                            ws.merge_range(row_num, 49, row_count, 49, participante.desempate_revisado_por.username if participante.desempate_revisado_por else '', formatoceldacenter)
                            ws.merge_range(row_num, 50, row_count, 50, participante.desempate_fecha_revision if participante.desempate_fecha_revision else '', formatoceldafecha)
                            for dato in titulos:
                                ws.write(row_num, 16, "{}".format(dato.titulo), formatoceldacenter)
                                ws.write(row_num, 17, "https://sga.unemi.edu.ec/{}".format(dato.archivo.url) if dato.archivo else '', formatoceldacenter)
                                row_num += 1
                        else:
                            ws.write(row_num, 0, participante.partida.periodoplanificacion.nombre, formatoceldacenter)
                            ws.write(row_num, 1, str(participante.partida), formatoceldacenter)
                            ws.write(row_num, 2, participante.partida.temadisertacion, formatoceldacenter)
                            ws.write(row_num, 3, participante.partida.carrera.__str__() if participante.partida.carrera else '', formatoceldacenter)
                            asignaturas = ", ".join([asig.__str__() for asig in participante.partida.partidas_asignaturas()])
                            ws.write(row_num, 4, asignaturas, formatoceldacenter)
                            ws.write(row_num, 5, participante.partida.get_dedicacion_display(), formatoceldacenter)
                            ws.write(row_num, 6, participante.partida.rmu, formatoceldacenter)
                            ws.write(row_num, 7, participante.partida.get_nivel_display(), formatoceldacenter)
                            ws.write(row_num, 8, participante.partida.get_modalidad_display(), formatoceldacenter)
                            ws.write(row_num, 9, participante.partida.get_jornada_display(), formatoceldacenter)
                            campo_amplio = ", ".join([camp.__str__() for camp in participante.partida.campoamplio.all()])
                            ws.write(row_num, 10, campo_amplio, formatoceldacenter)
                            campo_especifico = ", ".join([camp.__str__() for camp in participante.partida.campoespecifico.all()])
                            ws.write(row_num, 11, campo_especifico, formatoceldacenter)
                            campo_detallado = ", ".join([camp.__str__() for camp in participante.partida.campodetallado.all()])
                            ws.write(row_num, 12, campo_detallado, formatoceldacenter)
                            ws.write(row_num, 13, participante.id, formatoceldacenter)
                            ws.write(row_num, 14, f'{participante.persona.apellido1} {participante.persona.apellido2}', formatoceldacenter)
                            ws.write(row_num, 15, f'{participante.persona.nombres}', formatoceldacenter)
                            titulopersona = titulos.first()
                            archivotitulo = ''
                            nombre_titulo = ''
                            if titulopersona:
                                nombre_titulo = titulopersona.titulo.__str__()
                                archivotitulo = f"https://sga.unemi.edu.ec/{titulopersona.archivo.url}" if titulopersona.archivo  else ""
                            ws.write(row_num, 16, nombre_titulo, formatoceldacenter)
                            ws.write(row_num, 17, archivotitulo, formatoceldacenter)
                            ws.write(row_num, 18, participante.persona.cedula, formatoceldacenter)
                            ws.write(row_num, 19, participante.persona.email, formatoceldacenter)
                            ws.write(row_num, 20, participante.persona.telefono, formatoceldacenter)
                            ws.write(row_num, 21, participante.persona.telefono_conv, formatoceldacenter)
                            ws.write(row_num, 22, participante.get_estado_display(), formatoceldacenter)
                            ws.write(row_num, 23, 'SI' if participante.tiene_formacionacademica() else 'NO', formatoceldacenter)
                            ws.write(row_num, 24, 'SI' if participante.tiene_experienciapartida() else 'NO', formatoceldacenter)
                            ws.write(row_num, 25, 'SI' if participante.tiene_capacitaciones() else 'NO', formatoceldacenter)
                            ws.write(row_num, 26, 'SI' if participante.tiene_publicaciones() else 'NO', formatoceldacenter)
                            ws.write(row_num, 27, 'SI' if participante.tiene_idiomas() else 'NO', formatoceldacenter)
                            ws.write(row_num, 28, participante.pgradoacademico, formatoceldacenter)
                            ws.write(row_num, 29, participante.obsgradoacademico, formatoceldacenter)
                            ws.write(row_num, 30, participante.pexpdocente, formatoceldacenter)
                            ws.write(row_num, 31, participante.obsexperienciadoc, formatoceldacenter)
                            ws.write(row_num, 32, participante.pexpadministrativa, formatoceldacenter)
                            ws.write(row_num, 33, participante.obsexperienciaadmin, formatoceldacenter)
                            ws.write(row_num, 34, participante.pcapacitacion, formatoceldacenter)
                            ws.write(row_num, 35, participante.obscapacitacion, formatoceldacenter)
                            ws.write(row_num, 36, 'SI' if participante.aplico_desempate else 'NO', formatoceldacenter)
                            ws.write(row_num, 37, participante.nota_desempate, formatoceldacenter)
                            ws.write(row_num, 38, participante.nota_final_meritos, formatoceldacenter)
                            ws.write(row_num, 39, participante.obsgeneral, formatoceldacenter)
                            ws.write(row_num, 40, participante.get_estado_display(), formatoceldacenter)
                            ws.write(row_num, 41, 'SI' if participante.calificada else 'NO', formatoceldacenter)
                            ws.write(row_num, 42, 'SI' if participante.solapelacion else 'NO', formatoceldacenter)
                            participante_apelacion = participante.traer_apelacion()
                            ws.write(row_num, 43, participante_apelacion.get_estado_display() if participante_apelacion else "", formatoceldacenter)
                            obs_revisor, revisado_por, fecha_revision = '', '', ''
                            if participante_apelacion:
                                if participante_apelacion.estado != 0:
                                    obs_revisor = participante_apelacion.observacion_revisor
                                    revisado_por = participante_apelacion.revisado_por.username if participante_apelacion.revisado_por else ''
                                    fecha_revision = participante_apelacion.fecha_revision

                            ws.write(row_num, 44, obs_revisor, formatoceldacenter)
                            ws.write(row_num, 45, revisado_por, formatoceldacenter)
                            ws.write(row_num, 46, fecha_revision, formatoceldafecha)
                            ws.write(row_num, 47, participante.revisado_por.username if participante.revisado_por else '', formatoceldacenter)
                            ws.write(row_num, 48, participante.fecha_revision if participante.fecha_revision else '', formatoceldafecha)
                            ws.write(row_num, 49, participante.desempate_revisado_por.username if participante.desempate_revisado_por else '', formatoceldacenter)
                            ws.write(row_num, 50, participante.desempate_fecha_revision if participante.desempate_fecha_revision else '', formatoceldafecha)
                            row_num += 1
                    workbook.close()
                    output.seek(0)
                    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    response['Content-Disposition'] = f'attachment; filename="{name_file}"'
                    return response
                except Exception as ex:
                    pass

            if action == 'convocatoria':
                try:
                    data['title'] = u'Periodo de Planificación'
                    periodoacademico = PeriodoAcademicoConvocatoria.objects.get(id=int(encrypt(request.GET['id'])))
                    search, filtro, url_vars = request.GET.get('s', ''), (Q(status=True) & Q(vigente=True) & Q(periodo=periodoacademico)), ''

                    if search:
                        data['search'] = search
                        url_vars += "&s={}".format(search)
                        filtro = filtro & (Q(nombre__icontains=search))

                    listado = PeriodoPlanificacion.objects.filter(filtro).order_by('-id')
                    paging = MiPaginador(listado, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data["url_vars"] = url_vars
                    data['listado'] = page.object_list
                    data['list_count'] = len(listado)
                    return render(request, "postulate/adm_periodoplanificacion/view.html", data)
                except Exception as ex:
                    pass

            elif action == 'preguntasinformativas':
                try:
                    form = PreguntaPeriodoPlanificacionForm()
                    data['filtro'] = periodo_p = PeriodoAcademicoConvocatoria.objects.get(pk=int(encrypt(request.GET['id'])))
                    data['listado'] = periodo_p.preguntas()
                    data['form'] = form
                    template = get_template("postulate/adm_periodoplanificacion/modal/formpreguntas.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            if action == 'verhistorial':
                try:
                    data['title'] = u'Ver Historial'
                    data['id'] = id = encrypt(request.GET['id'])
                    data['filtro'] = filtro = Partida.objects.get(pk=int(id))
                    data['detalle'] = HistorialPartida.objects.filter(status=True, partida=filtro).order_by('-pk')
                    template = get_template("postulate/adm_periodoplanificacion/modal/historial.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

        else:
            try:
                data['title'] = u'Periodos'
                search, filtro, url_vars = request.GET.get('s', ''), ( Q(vigente=True) & Q(status=True)), ''

                if search:
                    data['search'] = search
                    url_vars += "&s={}".format(search)
                    filtro = filtro & (Q(nombre__icontains=search))

                listado = PeriodoAcademicoConvocatoria.objects.filter(filtro).order_by('-id')
                paging = MiPaginador(listado, 20)
                p = 1
                try:
                    paginasesion = 1
                    if 'paginador' in request.session:
                        paginasesion = int(request.session['paginador'])
                    if 'page' in request.GET:
                        p = int(request.GET['page'])
                    else:
                        p = paginasesion
                    try:
                        page = paging.page(p)
                    except:
                        p = 1
                    page = paging.page(p)
                except:
                    page = paging.page(p)
                request.session['paginador'] = p
                data['paging'] = paging
                data['rangospaging'] = paging.rangos_paginado(p)
                data['page'] = page
                data["url_vars"] = url_vars
                data['periodos'] = page.object_list
                data['list_count'] = len(listado)
                return render(request, "postulate/adm_periodoplanificacion/viewperiodo.html", data)
            except Exception as ex:
                pass


def fecha_letra(valor):
    mes = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
    a = int(valor[0:4])
    m = int(valor[5:7])
    d = int(valor[8:10])
    if d == 1:
        return u"al %s día del mes de %s" % (numero_a_letras(d), str(mes[m - 1]))
    else:
        return u"a los %s días del mes de %s" % (numero_a_letras(d), str(mes[m - 1]))