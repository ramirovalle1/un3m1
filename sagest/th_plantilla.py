# -*- coding: UTF-8 -*-
import json
import random
from datetime import datetime, timedelta

import os
import xlrd
from decimal import Decimal

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import Group
from django.db import transaction, connection
from django.db.models import Q
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.shortcuts import render
from django.template import RequestContext, Context
from django.template.loader import get_template
from openpyxl import load_workbook
from xlwt import *
from xlwt import easyxf
import xlwt

from decorators import secure_module
from sagest.forms import DistributivoPersonaForm
from sagest.models import DistributivoPersona, RegimenLaboral, NivelOcupacional, ModalidadLaboral, EstadoPuesto, \
    EscalaOcupacional, DenominacionPuesto, PuestoAdicional, Departamento, EstructuraProgramatica, \
    DistributivoPersonaHistorial
from settings import EMAIL_DOMAIN, ARCHIVO_TIPO_GENERAL, SEXO_MASCULINO, \
    EMAIL_INSTITUCIONAL_AUTOMATICO, PROFESORES_GROUP_ID, SITE_STORAGE
from sga.commonviews import adduserdata
from sga.excelbackground import masivo_datos_sistema_gobierno_background
from sga.forms import ImportarArchivoXLSForm
from sga.funciones import MiPaginador, log, generar_nombre, calculate_username, generar_usuario, variable_valor, \
    convertir_fecha
from sga.models import Archivo, Persona, Provincia, Canton, Administrativo, Profesor, Coordinacion, \
    TiempoDedicacionDocente, miinstitucion, CUENTAS_CORREOS, Titulacion, Notificacion
from sga.tasks import send_html_mail, conectar_cuenta
from settings import MEDIA_ROOT,MEDIA_URL
import io
import xlsxwriter
from urllib.request import urlopen


@login_required(redirect_field_name='ret', login_url='/loginsagest')
@secure_module
@transaction.atomic()
def view(request):
    data = {}
    adduserdata(request, data)
    persona = request.session['persona']
    usuario = request.user
    if request.method == 'POST':
        action = request.POST['action']

        if action == 'add':
            try:
                form = DistributivoPersonaForm(request.POST)
                if form.is_valid():
                    pers = form.cleaned_data['persona']
                    if DistributivoPersona.objects.filter(persona=pers,
                                                          regimenlaboral=form.cleaned_data['regimenlaboral'],
                                                          status=True).exists():
                        return JsonResponse({"result": "bad", "mensaje": "Registro Repetido."})

                    distributivopersona = DistributivoPersona(persona=pers,
                                                              regimenlaboral=form.cleaned_data['regimenlaboral'],
                                                              nivelocupacional=form.cleaned_data['nivelocupacional'],
                                                              modalidadlaboral=form.cleaned_data['modalidadlaboral'],
                                                              partidaindividual=form.cleaned_data['partidaindividual'],
                                                              estadopuesto=form.cleaned_data['estadopuesto'],
                                                              grado=form.cleaned_data['grado'],
                                                              rmuescala=Decimal(form.cleaned_data['rmuescala']).quantize(Decimal('.01')),
                                                              rmupuesto=Decimal(form.cleaned_data['rmupuesto']).quantize(Decimal('.01')),
                                                              rmusobrevalorado=Decimal(form.cleaned_data['rmupuesto']).quantize(Decimal('.01')),
                                                              escalaocupacional=form.cleaned_data['escalaocupacional'],
                                                              rucpatronal=form.cleaned_data['rucpatronal'],
                                                              codigosucursal=form.cleaned_data['codigosucursal'],
                                                              tipoidentificacion=form.cleaned_data['tipoidentificacion'],
                                                              denominacionpuesto=form.cleaned_data['denominacionpuesto'],
                                                              puestoadicinal=form.cleaned_data['puestoadicinal'],
                                                              unidadorganica=form.cleaned_data['unidadorganica'],
                                                              aporteindividual=Decimal(form.cleaned_data['aporteindividual']).quantize(Decimal('.01')),
                                                              aportepatronal=Decimal(form.cleaned_data['aportepatronal']).quantize(Decimal('.01')),
                                                              estructuraprogramatica=form.cleaned_data['estructuraprogramatica'],
                                                              comisioservicios=form.cleaned_data['comisioservicios'])
                    distributivopersona.save(request)
                    log(u'Nuevo distributivo personal: %s' % distributivopersona, request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": "Error al guardar los datos."})

        if action == 'edit':
            try:
                form = DistributivoPersonaForm(request.POST)
                form.fields['persona'].required = False
                if form.is_valid():
                    distributivopersona = DistributivoPersona.objects.filter(pk=int(request.POST['id']))[0]
                    if DistributivoPersona.objects.filter(persona=distributivopersona.persona,
                                                          regimenlaboral=form.cleaned_data['regimenlaboral'],
                                                          status=True).exclude(pk=int(request.POST['id'])).exists():
                        return JsonResponse({"result": "bad", "mensaje": "Registro Repetido."})

                    # distributivopersona.persona_id=form.cleaned_data['persona']
                    distributivopersona.regimenlaboral=form.cleaned_data['regimenlaboral']
                    distributivopersona.nivelocupacional=form.cleaned_data['nivelocupacional']
                    distributivopersona.modalidadlaboral=form.cleaned_data['modalidadlaboral']
                    distributivopersona.partidaindividual=form.cleaned_data['partidaindividual']
                    distributivopersona.estadopuesto=form.cleaned_data['estadopuesto']
                    distributivopersona.grado=form.cleaned_data['grado']
                    distributivopersona.rmuescala=Decimal(form.cleaned_data['rmuescala']).quantize(Decimal('.01'))
                    distributivopersona.rmupuesto=Decimal(form.cleaned_data['rmupuesto']).quantize(Decimal('.01'))
                    distributivopersona.rmusobrevalorado=Decimal(form.cleaned_data['rmupuesto']).quantize(Decimal('.01'))
                    distributivopersona.escalaocupacional=form.cleaned_data['escalaocupacional']
                    distributivopersona.rucpatronal=form.cleaned_data['rucpatronal']
                    distributivopersona.codigosucursal=form.cleaned_data['codigosucursal']
                    distributivopersona.tipoidentificacion=form.cleaned_data['tipoidentificacion']
                    distributivopersona.denominacionpuesto=form.cleaned_data['denominacionpuesto']
                    distributivopersona.puestoadicinal=form.cleaned_data['puestoadicinal']
                    distributivopersona.unidadorganica=form.cleaned_data['unidadorganica']
                    distributivopersona.aporteindividual=Decimal(form.cleaned_data['aporteindividual']).quantize(Decimal('.01'))
                    distributivopersona.aportepatronal=Decimal(form.cleaned_data['aportepatronal']).quantize(Decimal('.01'))
                    distributivopersona.estructuraprogramatica=form.cleaned_data['estructuraprogramatica']
                    distributivopersona.comisioservicios=form.cleaned_data['comisioservicios']
                    distributivopersona.save(request)
                    log(u'Modificacion distributivo personal: %s' % distributivopersona, request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": "Error al guardar los datos."})

        elif action == 'importar':
            try:
                form = ImportarArchivoXLSForm(request.POST, request.FILES)
                if form.is_valid():
                    hoy = datetime.now().date()
                    nfile = request.FILES['archivo']
                    data['namefile'] = nfile._name
                    numeroarchivo = 0
                    # try:
                    #     numeroarchivo = int(nfile._name.split('_')[1].split('(')[0].split('.')[0])
                    # except:
                    #     return HttpResponse(json.dumps({"result": "bad", "mensaje": u"Error de numeracion de archivo."}), content_type="application/json")
                    nfile._name = generar_nombre("importacion_", nfile._name)
                    # if DistributivoPersona.objects.filter(numeroarchivo__gte=numeroarchivo).exists():
                    #     return HttpResponse(json.dumps({"result": "bad", "mensaje": u"Error al subir un archivo antiguo."}), content_type="application/json")
                    archivo = Archivo(nombre='IMPORTACION TALENTO HUMANO',
                                      fecha=datetime.now().date(),
                                      archivo=nfile,
                                      tipo_id=ARCHIVO_TIPO_GENERAL)
                    archivo.save(request)
                    data['id_archivo']=archivo.id
                    data['numeroarchivo']=numeroarchivo
                    noti = Notificacion(cuerpo='Guardado de registros en progreso',
                                        titulo='Importación de datos del sistema de gobierno en proceso', destinatario=persona,
                                        url='',
                                        prioridad=1, app_label='SAGEST',
                                        fecha_hora_visible=datetime.now() + timedelta(days=1), tipo=2,
                                        en_proceso=True)
                    noti.save(request)
                    # workbook = xlrd.open_workbook(archivo.archivo.file.name)
                    messages.info(request, 'Importación de datos del sistema de gobierno en proceso, se le notificara cuando finalice la ejecución.')
                    masivo_datos_sistema_gobierno_background(request=request, data=data, notif=noti.pk).start()
                    log(u'Importo distributivo personal: %s' % persona, request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})

        if action == 'detalle':
            try:
                data = {}
                data['detalle'] = DistributivoPersona.objects.get(pk=int(request.POST['id']))
                template = get_template("th_plantilla/detalle.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'data': json_content})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        if action == 'lista_rol_excel':
            try:
                anio = request.POST['anio']
                directory = os.path.join(MEDIA_ROOT, 'reportes', 'rolpago')
                try:
                    os.stat(directory)
                except:
                    os.mkdir(directory)
                name_document = 'reporte_roles_pagos'
                nombre_archivo = name_document + "_{}.xlsx".format(random.randint(1, 10000).__str__())
                directory = os.path.join(MEDIA_ROOT, 'reportes', 'rolpago', nombre_archivo)


                __author__ = 'Unemi'

                workbook = xlsxwriter.Workbook(directory, {'constant_memory': True})
                ws = workbook.add_worksheet('exp_xls_post_part')
                style0 = workbook.add_format({'font_name':'Times New Roman'})
                style_nb = workbook.add_format({'font_name':'Times New Roman','bold':1})
                style_sb = workbook.add_format({'font_name':'Times New Roman','bold':1})
                title = workbook.add_format({
                        'bold': 1,
                        'border': 1,
                        'align': 'center',
                        'valign': 'vcenter',
                        'bg_color': 'silver',
                        'text_wrap': 1,
                        'font_name':'Times New Roman'})
                font_style = workbook.add_format({'font_name':'Times New Roman','bold':1})
                font_style2 = workbook.add_format({'font_name':'Times New Roman','bold':0})




                ws.merge_range(0, 0, 0, 10, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                # ws.write_merge(0, 0, 1, 10, 'DEPARTAMENTO FINANCIERO', title)
                # ws.write_merge(0, 0, 2, 10, 'SECCION CONTABILIDAD', title)
                # ws.write_merge(0, 0, 3, 10, 'PROYECCION DE GASTOS PERSONALES', title)
                # ws.write_merge(0, 0, 4, 10, anio, title)
                # response = HttpResponse(content_type="application/ms-excel")
                # response['Content-Disposition'] = 'attachment; filename=gastos_personales_' + random.randint(1, 10000).__str__() + '.xls'

                ws.set_column(0,0,73)
                ws.set_column(1,1,14)
                ws.set_column(2,2,60)
                ws.set_column(3,4,52)
                ws.set_column(5,5,20)
                ws.set_column(6,6,10)

                columns = [
                    (u"NOMINA", 10000),
                    (u"MES", 6000),
                    (u"DEPARTAMENTO", 8000),
                    (u"PERSONA", 6000),
                    (u"RUBRO", 6000),
                    (u"TIPORUBRO", 6000),
                    (u"VALOR", 6000),
                ]
                row_num = 5
                for col_num in range(len(columns)):
                    ws.write(row_num, col_num, columns[col_num][0], font_style)
                cursor = connection.cursor()
                sql = "select p.descripcion, (CASE WHEN p.mes=1 THEN 'ENERO' WHEN p.mes=2 THEN 'FEBRERO' WHEN p.mes=3 THEN 'MARZO' WHEN p.mes=4 THEN 'ABRIL' " \
                      " WHEN p.mes=5 THEN 'MAYO' WHEN p.mes=6 THEN 'JUNIO' WHEN p.mes=7 THEN 'JULIO' WHEN p.mes=8 THEN 'AGOSTO' WHEN p.mes=9 THEN 'SEPTIEMBRE' " \
                      " WHEN p.mes=10 THEN 'OCTUBRE' WHEN p.mes=11 THEN 'NOVIEMBRE' WHEN p.mes=12 THEN 'DICIEMBRE' end), de.nombre as departamento,pe.apellido1||' '||pe.apellido2||' '||pe.nombres as persona, " \
                      " r.descripcion as rubro, (case when r.tiporubro=1 then 'INGRESO'  when r.tiporubro=2 then 'EGRESO'  when r.tiporubro=3 then 'INFORMATIVO' end) as tiporubro, d.valor " \
                      " from sagest_periodorol p, sagest_detalleperiodorol d, sagest_rubrorol r, sga_persona pe, sagest_rolpago ro, sagest_departamento de " \
                      " where p.status=true and p.anio="+ anio +" and p.id=d.periodo_id and d.status=true and d.rubro_id=r.id and r.status=true and pe.id=d.persona_id and ro.periodo_id=d.periodo_id and " \
                                                                " d.persona_id=ro.persona_id and ro.status=true and ro.unidadorganica_id=de.id and de.status=true order by p.id, p.mes,de.nombre, pe.apellido1, pe.apellido2, pe.nombres, r.descripcion "

                cursor.execute(sql)
                results = cursor.fetchall()
                row_num = 6
                for r in results:
                    i = 0
                    campo1 = r[0]
                    campo2 = r[1]
                    campo3 = r[2]
                    campo4 = r[3]
                    campo5 = r[4]
                    campo6 = r[5]
                    campo7 = r[6]

                    ws.write(row_num, 0, campo1, font_style2)
                    ws.write(row_num, 1, campo2, font_style2)
                    ws.write(row_num, 2, campo3, font_style2)
                    ws.write(row_num, 3, campo4, font_style2)
                    ws.write(row_num, 4, campo5, font_style2)
                    ws.write(row_num, 5, campo6, font_style2)
                    ws.write(row_num, 6, campo7, font_style2)

                    row_num += 1
                workbook.close()
                response = HttpResponse(directory, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = 'attachment; filename=%s' % name_document
                url_file='{}reportes/rolpago/{}'.format(MEDIA_URL,nombre_archivo)
                return JsonResponse({'result':'ok','archivo':url_file})
            except Exception as ex:
                pass

        if action == 'descargarhistoricofecha':
            try:
                __author__ = 'Unemi'
                fechai = convertir_fecha(request.POST['fechai'])
                fechaf = convertir_fecha(request.POST['fechaf'])
                style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
                style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
                style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
                title = easyxf(
                    'font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                style1 = easyxf(num_format_str='D-MMM-YY')
                output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media', 'rolpago'))
                nombre = "DISTRIBUTIVO_" + datetime.now().strftime('%Y%m%d_%H%M%S') + ".xls"
                ruta = "media/rolpago/" + nombre
                filename = os.path.join(output_folder, nombre)
                font_style = XFStyle()
                font_style.font.bold = True
                font_style2 = XFStyle()
                font_style2.font.bold = False
                wb = Workbook(encoding='utf-8')
                ws = wb.add_sheet('exp_xls_post_part')
                ws.write_merge(0, 0, 0, 7, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                response = HttpResponse(content_type="application/ms-excel")
                response['Content-Disposition'] = 'attachment; filename=empleados_' + random.randint(1,10000).__str__() + '.xls'
                row_num = 3
                columns = [
                    (u"REGIMEN LABORAL", 12000),
                    (u"FECHA HISTORIAL", 6000),
                    (u"NIVEL OCUPACIONAL", 6000),
                    (u"MODALIDAD LABORAL", 6000),
                    (u"NÚMERO IDENTIFICACIÓN", 6000),
                    (u"NOMBRES", 6000),
                    (u"DENOMINACIÓN PUESTO", 12000),
                    (u"ACUMULACIÓN DÉCIMO TERCERO", 6000),
                    (u"ACUMULACIÓN DÉCIMO CUARTO", 6000),
                    (u"RMU PUESTO", 6000),
                    (u"FONDO RESERVA", 6000),
                    (u"TOTAL INGRESOS", 6000),
                    (u"DÉCIMO TERCERO", 6000),
                    (u"DÉCIMO CUARTO", 6000),
                    (u"TOTAL INGRESO", 6000),
                    (u"OBSERVACIONES", 6000),
                ]
                for col_num in range(len(columns)):
                    ws.write(row_num, col_num, columns[col_num][0], font_style)
                    ws.col(col_num).width = columns[col_num][1]
                # cursor = connection.cursor()
                date_format = xlwt.XFStyle()
                date_format.num_format_str = 'yyyy/mm/dd'
                listadodistributivo = DistributivoPersonaHistorial.objects.values_list('regimenlaboral__descripcion', 'fechahistorial', 'nivelocupacional__descripcion', 'modalidadlaboral__descripcion', 'persona__cedula', 'persona__apellido1','persona__apellido2', 'persona__nombres', 'denominacionpuesto__descripcion', 'rmupuesto').filter(status=True, fechahistorial__range=(fechai,fechaf)).distinct().order_by('regimenlaboral__descripcion','persona__cedula')
                row_num = 4
                for lista in listadodistributivo:
                    i = 0
                    campo1 = lista[0]
                    campo2 = f"{lista[1].date()}" if lista[1] else ''
                    campo3 = lista[2]
                    campo4 = lista[3]
                    campo5 = lista[4]
                    campo6 = u"%s %s %s" % (lista[5], lista[6], lista[7])
                    campo7 = lista[8]
                    campo8 = ''
                    campo9 = ''
                    campo10 = lista[9]
                    campo11 = 0
                    campo12 = 0
                    campo13 = 0
                    campo14 = 0
                    campo15 = 0
                    campo16 = ''

                    ws.write(row_num, 0, campo1, font_style2)
                    ws.write(row_num, 1, campo2, font_style2)
                    ws.write(row_num, 2, campo3, font_style2)
                    ws.write(row_num, 3, campo4, font_style2)
                    ws.write(row_num, 4, campo5, font_style2)
                    ws.write(row_num, 5, campo6, font_style2)
                    ws.write(row_num, 6, campo7, font_style2)
                    ws.write(row_num, 7, campo8, font_style2)
                    ws.write(row_num, 8, campo9, font_style2)
                    ws.write(row_num, 9, campo10, font_style2)
                    ws.write(row_num, 10, campo11, font_style2)
                    ws.write(row_num, 11, campo12, font_style2)
                    ws.write(row_num, 12, campo13, font_style2)
                    ws.write(row_num, 13, campo14, font_style2)
                    ws.write(row_num, 14, campo15, font_style2)
                    ws.write(row_num, 15, campo15, font_style2)

                    row_num += 1
                wb.save(filename)
                # connection.close()
                return JsonResponse({'result': 'ok', 'archivo': ruta})
            except Exception as ex:
                pass

        if action == 'cargartabla':
            try:
                with transaction.atomic():
                    if not 'archivo' in request.FILES:
                        transaction.set_rollback(True)
                        return JsonResponse({"result": "bad", "mensaje": u"Por Favor seleccione un archivo."})
                    archivo=request.FILES['archivo']
                    workbook = load_workbook(archivo)
                    sheet = workbook.active
                    datos = sheet.iter_rows()
                    linea = 1
                    personas=[]
                    departamentos=[]
                    for rowx in datos:
                        if linea >= 2:
                            cols = [str(cell.value) for cell in rowx]
                            identificacion = str(cols[20]).strip().upper()
                            if not Departamento.objects.filter(nombre=str(cols[30]).strip().upper()).exists():
                                if not str(cols[30]).strip() in departamentos:
                                    departamentos.append(str(cols[30]).strip())

                            if not Persona.objects.filter(Q(cedula=identificacion) | Q(pasaporte=identificacion)).exists():
                                diccionario={'cedula':str(cols[20]).strip(),
                                             'nombres':str(cols[21]).strip(),
                                             }
                                personas.append(diccionario)
                        linea+=1
                    data['personas']=personas
                    data['cantpersonas']=len(personas)
                    data['departamentos']=departamentos
                    data['cantdepa'] = len(departamentos)
                    data['nombrearchivo']=archivo._name
                    template = get_template("th_plantilla/detalletabla.html")
                    return JsonResponse({"result":"ok", "data":template.render(data)})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al mostrar los datos."})

        return JsonResponse({"result": "bad", "mensaje": u"Solicitud Incorrecta."})
    else:
        if 'action' in request.GET:
            action = request.GET['action']

            if action == 'importar':
                try:
                    data['title'] = u'Importar datos del sistema de gobierno'
                    data['form'] = ImportarArchivoXLSForm()
                    return render(request, "th_plantilla/importar.html", data)
                except Exception as ex:
                    pass

            if action == 'add':
                try:
                    data['title'] = u'Adicionar distributivo del personal'
                    form = DistributivoPersonaForm()
                    form.fields['persona'].queryset = Persona.objects.none()
                    data['form'] = form
                    return render(request, "th_plantilla/add.html", data)
                except:
                    pass

            if action == 'edit':
                try:
                    data['title'] = u'Editar distributivo Personal'
                    data['distributivopersona'] = distributivopersona = DistributivoPersona.objects.filter(pk=request.GET['id'])[0]
                    form = DistributivoPersonaForm(initial={'regimenlaboral': distributivopersona.regimenlaboral,
                                                            'nivelocupacional': distributivopersona.nivelocupacional,
                                                            'modalidadlaboral': distributivopersona.modalidadlaboral,
                                                            'partidaindividual': distributivopersona.partidaindividual,
                                                            'estadopuesto': distributivopersona.estadopuesto,
                                                            'grado': distributivopersona.grado,
                                                            'rmuescala': distributivopersona.rmuescala,
                                                            'rmupuesto': distributivopersona.rmupuesto,
                                                            'rmusobrevalorado': distributivopersona.rmusobrevalorado,
                                                            'escalaocupacional': distributivopersona.escalaocupacional,
                                                            'rucpatronal': distributivopersona.rucpatronal,
                                                            'codigosucursal': distributivopersona.codigosucursal,
                                                            'tipoidentificacion': distributivopersona.tipoidentificacion,
                                                            'denominacionpuesto': distributivopersona.denominacionpuesto,
                                                            'puestoadicinal': distributivopersona.puestoadicinal,
                                                            'unidadorganica': distributivopersona.unidadorganica,
                                                            'aporteindividual': distributivopersona.aporteindividual,
                                                            'aportepatronal': distributivopersona.aportepatronal,
                                                            'estructuraprogramatica': distributivopersona.estructuraprogramatica,
                                                            'comisioservicios': distributivopersona.comisioservicios})
                    form.edit(distributivopersona)
                    form.fields['persona'].queryset = Persona.objects.filter(id=distributivopersona.persona.id)
                    data['form'] = form
                    return render(request, "th_plantilla/edit.html", data)
                except:
                    pass

            if action == 'descargar':
                try:
                    __author__ = 'Unemi'
                    style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
                    style_nb = easyxf('font: name Times New Roman, color-index blue, bold on',   num_format_str='#,##0.00')
                    style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
                    title = easyxf(
                        'font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                    style1 = easyxf(num_format_str='D-MMM-YY')
                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('exp_xls_post_part')
                    ws.write_merge(0, 0, 0, 7, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=empleados_' + random.randint(1, 10000).__str__() + '.xls'
                    row_num = 3
                    columns = [
                        (u"CÓDIGO BIOMETRICO", 6000),
                        (u"DEPARTAMENTO", 6000),
                        (u"CEDULA", 6000),
                        (u"APELLIDO 1", 6000),
                        (u"APELLIDO 2", 6000),
                        (u"NOMBRE", 6000),
                        (u"FECHA NACIMIENTO", 6000),
                        (u"EMAIL INSTITUCIONAL", 6000),
                        (u"REGIMEN LABORAL", 12000),
                        (u"DENOMINACIÓN PUESTO", 12000),
                        (u"NIVEL OCUPACIONAL", 6000),
                        (u"MODALIDAD LABORAL", 6000),
                        (u"USUARIO", 6000),
                        (u"SEXO", 6000),
                        (u"RMU", 6000),
                        (u"TELEFONO MOVIL", 12000),
                        (u"EMAIL", 6000),
                        (u"TITULO", 6000),
                        (u"ETNIA", 4000),
                        (u"NACIONALIDAD", 4000),
                        (u"ESTADO CIVIL", 4000),
                        (u"DIRECCIÓN", 4000),
                        (u"CONTACTO DE EMERGENCIA (PARENTESCO: TELEFONO)", 4000),
                        (u"PAIS NACIMIENTO", 8000),
                        (u"PROVINCIA NACIMIENTO", 8000),
                        (u"CIUDAD NACIMIENTO", 8000),
                        (u"PAIS RESIDENCIA", 8000),
                        (u"PROVINCIA RESIDENCIA", 8000),
                        (u"CIUDAD RESIDENCIA", 8000),
                        (u"CALLE PRINCIPAL", 8000),
                        (u"CALLE SECUNDARIA", 8000),
                        (u"NUMERO  DOMICILIO", 8000),
                        (u"REFERENCIA", 8000),
                        (u"PARROQUIA DOMICILIO", 8000),
                        (u"SECTOR", 8000),
                        (u"TELÉFONO CONVENCIONAL", 8000),
                        (u"CODPERSONA", 6000),
                        (u"FOTO", 6000),
                    ]
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], font_style)
                        ws.col(col_num).width = columns[col_num][1]
                    cursor = connection.cursor()
                    date_format = xlwt.XFStyle()
                    date_format.num_format_str = 'yyyy/mm/dd'
                    if int(request.GET['s']) == 1:
                        listadodistributivo = DistributivoPersona.objects.filter(status=True).order_by('unidadorganica__nombre','regimenlaboral__id')
                    else:
                        listadodistributivo = DistributivoPersona.objects.filter(estadopuesto__id=1, status=True).order_by('unidadorganica__nombre', 'regimenlaboral__id')
                    row_num = 4
                    for lista in listadodistributivo:
                        i = 0
                        if (lista.persona != None):
                            campo1 = lista.persona.identificacioninstitucion
                            campo2 = lista.unidadorganica.nombre
                            campo3 = lista.persona.cedula
                            campo4 = lista.persona.apellido1
                            campo5 = lista.persona.apellido2
                            campo6 = lista.persona.nombres
                            campo7 = lista.persona.nacimiento
                            campo8 = lista.persona.emailinst
                            campo9 = lista.regimenlaboral.descripcion
                            campo10 = lista.denominacionpuesto.descripcion
                            campo11 = lista.nivelocupacional.descripcion
                            campo12 = lista.modalidadlaboral.descripcion
                            campo13 = lista.persona.usuario.username
                            campo14 = lista.persona.sexo.nombre
                            campo15 = lista.rmupuesto
                            campo16 = lista.persona.telefono
                            campo17 = lista.persona.email
                            campo18 = ''
                            campo19 = ''
                            campo21 = ''
                            if lista.persona.titulacion_set.filter(titulo__isnull=False, status=True, cursando=False):
                                campo18 = lista.persona.titulacion_set.filter(status=True, cursando=False).order_by("-titulo__nivel__rango")[0].titulo.nombre
                            if lista.persona.tiene_raza():
                                campo19 = lista.persona.mi_perfil().raza.nombre
                            campo20 = lista.persona.nacionalidad
                            if lista.persona.personaextension_set.values("id").exists():
                                if lista.persona.personaextension_set.filter(status=True)[0].estadocivil:
                                    campo21 = lista.persona.estado_civil().nombre
                            campo22 = lista.persona.direccion + ' ' + lista.persona.direccion2 + ' ' + lista.persona.referencia
                            campo23 = ''
                            campo24 = ''
                            campo25 = ''
                            campo26 = ''
                            campo27 = ''
                            campo28 = ''
                            if lista.persona.paisnacimiento_id:
                                campo23 = lista.persona.paisnacimiento.nombre
                            if lista.persona.provincianacimiento_id:
                                campo24 = lista.persona.provincianacimiento.nombre
                            if lista.persona.cantonnacimiento_id:
                                campo25 = lista.persona.cantonnacimiento.nombre
                            if lista.persona.pais_id:
                                campo26 = lista.persona.pais.nombre
                            if lista.persona.provincia_id:
                                campo27 = lista.persona.provincia.nombre
                            if lista.persona.canton_id:
                                campo28 = lista.persona.canton.nombre

                            ws.write(row_num, 0, campo1, font_style2)
                            ws.write(row_num, 1, campo2, font_style2)
                            ws.write(row_num, 2, campo3, font_style2)
                            ws.write(row_num, 3, campo4, font_style2)
                            ws.write(row_num, 4, campo5, font_style2)
                            ws.write(row_num, 5, campo6, style1)
                            ws.write(row_num, 6, campo7, date_format)
                            ws.write(row_num, 7, campo8, font_style2)
                            ws.write(row_num, 8, campo9, font_style2)
                            ws.write(row_num, 9, campo10, font_style2)
                            ws.write(row_num, 10, campo11, font_style2)
                            ws.write(row_num, 11, campo12, font_style2)
                            ws.write(row_num, 12, campo13, font_style2)
                            ws.write(row_num, 13, campo14, font_style2)
                            ws.write(row_num, 14, campo15, font_style2)
                            ws.write(row_num, 15, campo16, font_style2)
                            ws.write(row_num, 16, campo17, font_style2)
                            ws.write(row_num, 17, campo18, font_style2)
                            ws.write(row_num, 18, campo19, font_style2)
                            ws.write(row_num, 19, campo20, font_style2)
                            ws.write(row_num, 20, campo21, font_style2)
                            ws.write(row_num, 21, campo22, font_style2)
                            ws.write(row_num, 22, lista.persona.registro_emergencia(), font_style2)
                            ws.write(row_num, 23, campo23, font_style2)
                            ws.write(row_num, 24, campo24, font_style2)
                            ws.write(row_num, 25, campo25, font_style2)
                            ws.write(row_num, 26, campo26, font_style2)
                            ws.write(row_num, 27, campo27, font_style2)
                            ws.write(row_num, 28, campo28, font_style2)
                            ws.write(row_num, 29, u'%s' % lista.persona.direccion, font_style2)
                            ws.write(row_num, 30, u'%s' % lista.persona.direccion2, font_style2)
                            ws.write(row_num, 31, u'%s' % lista.persona.num_direccion, font_style2)
                            ws.write(row_num, 32, u'%s' % lista.persona.referencia, font_style2)
                            ws.write(row_num, 33, u'%s' % lista.persona.parroquia if lista.persona.parroquia else '', font_style2)
                            ws.write(row_num, 34, u'%s' % lista.persona.sector, font_style2)
                            ws.write(row_num, 35, u'%s' % lista.persona.telefono_conv, font_style2)
                            ws.write(row_num, 36, u'%s' % lista.persona.id, font_style2)
                            ws.write(row_num, 37, 'SI' if lista.persona.foto() else 'NO', font_style2)

                            row_num += 1
                    wb.save(response)
                    connection.close()
                    return response
                except Exception as ex:
                    pass

            if action == 'descargarfechasseparadas':
                try:
                    __author__ = 'Unemi'
                    style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
                    style_nb = easyxf('font: name Times New Roman, color-index blue, bold on',   num_format_str='#,##0.00')
                    style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
                    title = easyxf(
                        'font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                    style1 = easyxf(num_format_str='D-MMM-YY')
                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('exp_xls_post_part')
                    ws.write_merge(0, 0, 0, 7, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=empleados_fechasseparadas' + random.randint(1, 10000).__str__() + '.xls'
                    row_num = 3
                    columns = [
                        (u"DEPARTAMENTO", 6000),
                        (u"CEDULA", 6000),
                        (u"APELLIDO 1", 6000),
                        (u"APELLIDO 2", 6000),
                        (u"NOMBRE", 6000),
                        (u"FECHA NACIMIENTO", 6000),
                        (u"DIA", 6000),
                        (u"MES", 6000),
                        (u"ANIO", 6000),
                        (u"EMAIL INSTITUCIONAL", 6000),
                        (u"REGIMEN LABORAL", 12000),
                        (u"DENOMINACIÓN PUESTO", 12000),
                        (u"SEXO", 6000),
                        (u"TELEFONO MOVIL", 12000),
                        (u"EMAIL", 6000),
                        (u"TITULO", 6000),
                    ]
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], font_style)
                        ws.col(col_num).width = columns[col_num][1]
                    cursor = connection.cursor()
                    date_format = xlwt.XFStyle()
                    date_format.num_format_str = 'yyyy/mm/dd'
                    if int(request.GET['s']) == 1:
                        listadodistributivo = DistributivoPersona.objects.filter(status=True).order_by('unidadorganica__nombre','regimenlaboral__id')
                    else:
                        listadodistributivo = DistributivoPersona.objects.filter(estadopuesto__id=1, status=True).order_by('unidadorganica__nombre', 'regimenlaboral__id')
                    row_num = 4
                    for lista in listadodistributivo:
                        i = 0
                        if (lista.persona != None):
                            campo2 = lista.unidadorganica.nombre
                            campo3 = lista.persona.cedula
                            campo4 = lista.persona.apellido1
                            campo5 = lista.persona.apellido2
                            campo6 = lista.persona.nombres
                            campo7 = lista.persona.nacimiento
                            campo7dia = lista.persona.nacimiento.day
                            campo7mes = lista.persona.nacimiento.month
                            campo7anio = lista.persona.nacimiento.year
                            campo8 = lista.persona.emailinst
                            campo9 = lista.regimenlaboral.descripcion
                            campo10 = lista.denominacionpuesto.descripcion
                            campo11 = lista.persona.sexo.nombre
                            campo12 = lista.persona.telefono
                            campo13 = lista.persona.email
                            campo14 = ''
                            if lista.persona.titulacion_set.filter(titulo__isnull=False, status=True):
                                campo14 = lista.persona.titulacion_set.filter(status=True)[0].titulo.nombre
                            ws.write(row_num, 0, campo2, font_style2)
                            ws.write(row_num, 1, campo3, font_style2)
                            ws.write(row_num, 2, campo4, font_style2)
                            ws.write(row_num, 3, campo5, font_style2)
                            ws.write(row_num, 4, campo6, style1)
                            ws.write(row_num, 5, campo7, date_format)
                            ws.write(row_num, 6, campo7dia, font_style2)
                            ws.write(row_num, 7, campo7mes, font_style2)
                            ws.write(row_num, 8, campo7anio, font_style2)
                            ws.write(row_num, 9, campo8, font_style2)
                            ws.write(row_num, 10, campo9, font_style2)
                            ws.write(row_num, 11, campo10, font_style2)
                            ws.write(row_num, 12, campo11, font_style2)
                            ws.write(row_num, 13, campo12, font_style2)
                            ws.write(row_num, 14, campo13, font_style2)
                            ws.write(row_num, 15, campo14, font_style2)

                            row_num += 1
                    wb.save(response)
                    connection.close()
                    return response
                except Exception as ex:
                    pass

            if action == 'descargardocentes':
                try:
                    __author__ = 'Unemi'
                    style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
                    style_nb = easyxf('font: name Times New Roman, color-index blue, bold on',
                                      num_format_str='#,##0.00')
                    style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
                    title = easyxf(
                        'font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                    style1 = easyxf(num_format_str='D-MMM-YY')
                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('exp_xls_docentes_activos')
                    response = HttpResponse(content_type="application/ms-excel")
                    response[
                        'Content-Disposition'] = 'attachment; filename=exp_xls_docentes_activos_' + random.randint(1, 10000).__str__() + '.xls'
                    columns = [
                        (u"DEPARTAMENTO", 6000),
                        (u"FACULTAD", 6000),
                        (u"CEDULA", 6000),
                        (u"APELLIDO 1", 6000),
                        (u"APELLIDO 2", 6000),
                        (u"NOMBRE", 6000),
                        (u"FECHA NACIMIENTO", 6000),
                        (u"EMAIL INSTITUCIONAL", 6000),
                        (u"REGIMEN LABORAL", 12000),
                        (u"DENOMINACIÓN PUESTO", 12000),
                        (u"NIVEL OCUPACIONAL", 6000),
                        (u"MODALIDAD LABORAL", 6000),
                        (u"USUARIO", 0),
                        (u"SEXO", 6000),
                        (u"DEDICACION", 12000),
                        (u"TIPO PROFESOR", 12000),
                        (u"CATEGORIA", 12000),
                        (u"ESCALAFON", 12000),
                        (u"RMU", 6000),
                        (u"CELULAR", 6000),
                        (u"ETNIA", 6000),
                        (u"NACIONALIDAD", 4000),
                        (u"DIRECCIÓN", 4000),
                        (u"PROVÍNCIA", 4000),
                        (u"CIUDAD", 4000),
                        (u"CANTÓN", 4000),
                        (u"EMAIL PERSONAL", 6000),
                    ]
                    row_num = 0
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], font_style)
                        ws.col(col_num).width = columns[col_num][1]
                    cursor = connection.cursor()
                    # lista_json = []
                    # data = {}
                    sql = "SELECT dep.nombre, p.cedula, p.apellido1, p.apellido2, p.nombres, p.nacimiento, (usua.username || '@unemi.edu.ec') as emailinst , " \
                          " r.descripcion, de.descripcion, nio.descripcion, mod.descripcion, usua.username," \
                          "  (CASE p.sexo_id WHEN 1 THEN 'FEMENINO' ELSE 'MASCULINO' END) AS sexo," \
                          "  dp.nombre, (select cat.nombre from sga_categorizaciondocente cat where cat.id=pro.categoria_id)," \
                          " (select pti.nombre from sga_profesortipo pti where pti.id=pro.nivelcategoria_id)," \
                          "    (select nie.nombre from sga_nivelescalafondocente nie where nie.id=pro.nivelescalafon_id), d.rmupuesto,p.telefono,(select raz.nombre from sga_perfilinscripcion perfil,sga_raza raz where perfil.raza_id=raz.id and perfil.persona_id=p.id limit 1) as etnia,p.nacionalidad, " \
                          " ( select c1.nombre from sga_coordinacion c1 where c1.id=pro.coordinacion_id) as coordinacion_profesor, p.direccion || ' ' || p.direccion2 as direccion, " \
                          " prov.nombre, p.ciudad, cant.nombre,  p.email" \
                          "  from sga_persona p, auth_user usua, " \
                          "  sagest_distributivopersona d, sagest_nivelocupacional nio," \
                          "  sagest_modalidadlaboral mod, sagest_regimenlaboral r, sga_canton cant, " \
                          "  sagest_denominacionpuesto de, sagest_departamento dep , sga_profesor pro, sga_tiempodedicaciondocente dp, sga_provincia prov" \
                          "  where p.id=d.persona_id and d.status=true and d.estadopuesto_id=1 " \
                          "  and r.id=d.regimenlaboral_id and d.regimenlaboral_id=2 and pro.persona_id=p.id and pro.activo= true" \
                          "  and dp.id=pro.dedicacion_id and prov.id=p.provincia_id and cant.id = p.canton_id" \
                          "  and usua.id=p.usuario_id and d.nivelocupacional_id=nio.id and d.modalidadlaboral_id=mod.id and de.id=d.denominacionpuesto_id and dep.id=d.unidadorganica_id order by dep.nombre, r.id "

                    cursor.execute(sql)
                    results = cursor.fetchall()
                    row_num = 1
                    for r in results:
                        i = 0
                        campo1 = r[0]
                        campo2 = r[1]
                        campo3 = r[2]
                        campo4 = r[3]
                        campo5 = r[4]
                        campo6 = r[5]
                        campo7 = r[6]
                        campo8 = r[7]
                        campo9 = r[8]
                        campo10 = r[9]
                        campo11 = r[10]
                        campo12 = r[11]
                        campo13 = r[12]
                        if r[13]:
                            campo14 = r[13]
                        else:
                            campo14 = ''
                        if r[14]:
                            campo15 = r[14]
                        else:
                            campo15 = ''
                        if r[15]:
                            campo16 = "PROFESOR " + r[15]
                        else:
                            campo16 = ''
                        if r[16]:
                            campo17 = "PROFESOR " + r[16]
                        else:
                            campo17 = ''
                        campo18 = r[17]
                        campo19 = r[18]
                        campo20 = r[19]
                        campo21 = r[20]
                        campo22 = r[21]
                        campo23 = r[22]
                        campo24 = r[23]
                        campo25 = r[24]
                        campo26 = r[25]
                        campo27 = r[26]

                        ws.write(row_num, 0, campo1, font_style2)
                        ws.write(row_num, 1, campo22, font_style2)
                        ws.write(row_num, 2, campo2, font_style2)
                        ws.write(row_num, 3, campo3, font_style2)
                        ws.write(row_num, 4, campo4, font_style2)
                        ws.write(row_num, 5, campo5, font_style2)
                        ws.write(row_num, 6, campo6, style1)
                        ws.write(row_num, 7, campo7, font_style2)
                        ws.write(row_num, 8, campo8, font_style2)
                        ws.write(row_num, 9, campo9, font_style2)
                        ws.write(row_num, 10, campo10, font_style2)
                        ws.write(row_num, 11, campo11, font_style2)
                        ws.write(row_num, 12, campo12, font_style2)
                        ws.write(row_num, 13, campo13, font_style2)
                        ws.write(row_num, 14, campo14, font_style2)
                        ws.write(row_num, 15, campo16, font_style2)
                        ws.write(row_num, 16, campo15, font_style2)
                        ws.write(row_num, 17, campo17, font_style2)
                        ws.write(row_num, 18, campo18, font_style2)
                        ws.write(row_num, 19, campo19, font_style2)
                        ws.write(row_num, 20, campo20, font_style2)
                        ws.write(row_num, 21, campo21, font_style2)
                        ws.write(row_num, 22, campo23, font_style2)
                        ws.write(row_num, 23, campo24, font_style2)
                        ws.write(row_num, 24, campo25, font_style2)
                        ws.write(row_num, 25, campo26, font_style2)
                        ws.write(row_num, 26, campo27, font_style2)
                        row_num += 1
                    wb.save(response)
                    connection.close()
                    return response
                except Exception as ex:
                    pass

            if action == 'descargarforadmactivos':
                try:
                    tipotrabajador = request.GET['tipo']
                    cursor = connection.cursor()
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=listado activos.xls'
                    wb = xlwt.Workbook()
                    ws = wb.add_sheet('Sheetname')
                    estilo = xlwt.easyxf('font: height 350, name Arial, colour_index black, bold on, italic on; align: wrap on, vert centre, horiz center;')
                    ws.write_merge(0, 0, 0, 9, 'UNIVERSIDAD ESTATAL DE MILAGRO', estilo)
                    ws.col(0).width = 1000
                    ws.col(1).width = 3000
                    ws.col(2).width = 10000
                    ws.col(3).width = 5000
                    ws.col(4).width = 4000
                    ws.col(5).width = 10000
                    ws.col(6).width = 4000
                    ws.col(7).width = 10000
                    ws.col(8).width = 10000
                    ws.col(9).width = 6000
                    ws.write(4, 0, 'N.')
                    ws.write(4, 1, 'CEDULA')
                    ws.write(4, 2, 'NOMBRES')
                    ws.write(4, 3, 'SEXO')
                    ws.write(4, 4, 'FECHATITULO')
                    ws.write(4, 5, 'TITULO')
                    ws.write(4, 6, 'NIVEL')
                    ws.write(4, 7, 'GRADOTITULO')
                    ws.write(4, 8, 'INSTITUCION')
                    ws.write(4, 9, 'VERIFICADO')
                    ws.write(4, 10, 'UNIDAD')
                    ws.write(4, 11, 'CARGO')
                    ws.write(4, 12, 'RMU')
                    date_format = xlwt.XFStyle()
                    date_format.num_format_str = 'yyyy/mm/dd'
                    sql = "select distinct p.cedula,p.apellido1,p.apellido2,p.nombres," \
                          "d.regimenlaboral_id,ti.fechaobtencion,tit.nombre as titulo,nti.nombre as nivel," \
                          "gti.nombre as gradotitulo, sup.nombre as institucion,ti.verificado,sex.nombre as sexo, " \
                          "depa.nombre AS direccion, d.rmupuesto,deno.descripcion from sagest_distributivopersona d " \
                          "right join  sga_persona p on p.id=d.persona_id " \
                          "left join sga_sexo sex on p.sexo_id=sex.id " \
                          "left join sga_titulacion ti on ti.persona_id=p.id " \
                          "left join sga_titulo tit on tit.id=ti.titulo_id " \
                          "left join sga_niveltitulacion nti on nti.id=tit.nivel_id " \
                          "left join sga_gradotitulacion gti on gti.id=tit.grado_id " \
                          "left JOIN sagest_departamento depa ON depa.id=d.unidadorganica_id " \
                          "left JOIN sagest_denominacionpuesto deno ON deno.id=d.denominacionpuesto_id " \
                          "left join sga_institucioneducacionsuperior sup on sup.id = ti.institucion_id  " \
                          "where  d.regimenlaboral_id = '" + tipotrabajador + "' " \
                                                                              "and p.cedula IS NOT NULL " \
                                                                              "order by p.apellido1,p.apellido2,p.nombres "

                    cursor.execute(sql)
                    results = cursor.fetchall()
                    a = 4
                    for per in results:
                        a += 1
                        ws.write(a, 0, a - 4)
                        ws.write(a, 1, per[0])
                        ws.write(a, 2, per[1] + ' ' + per[2] + ' ' + per[3])
                        ws.write(a, 3, per[11])
                        ws.write(a, 4, per[5], date_format)
                        ws.write(a, 5, per[6])
                        ws.write(a, 6, per[7])
                        ws.write(a, 7, per[8])
                        ws.write(a, 8, per[9])
                        ws.write(a, 9, per[10])
                        ws.write(a, 10, per[12])
                        ws.write(a, 11, per[13])
                        ws.write(a, 12, per[14])
                    wb.save(response)
                    connection.close()
                    return response
                except Exception as ex:
                    pass

            if action == 'descargarlotaip':
                try:
                    __author__ = 'Unemi'
                    style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
                    style_nb = easyxf('font: name Times New Roman, color-index blue, bold on',
                                      num_format_str='#,##0.00')
                    style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
                    title = easyxf(
                        'font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                    style1 = easyxf(num_format_str='D-MMM-YY')
                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('exp_xls_post_part')
                    ws.write_merge(0, 0, 0, 7, 'UNIVERSIDAD ESTATAL DE MILAGRO', title)
                    response = HttpResponse(content_type="application/ms-excel")
                    response[
                        'Content-Disposition'] = 'attachment; filename=empleados_' + random.randint(
                        1, 10000).__str__() + '.xls'
                    columns = [
                        (u"N", 1000),
                        # (u"CODIGOPERSONA", 2000),
                        (u"CEDULA", 3000),
                        (u"NOMBRES", 11000),
                        (u"PUESTO", 11000),
                        (u"UNIDAD", 14444),
                        (u"DIRECCION INSTITUCIONAL", 9000),
                        (u"CIUDAD QUE LABORA", 5500),
                        (u"TELEFONO", 6000),
                        (u"EXTENSION", 3000),
                        (u"EMAIL INSTITUCIONAL", 7000),
                    ]
                    row_num = 3
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], font_style)
                        ws.col(col_num).width = columns[col_num][1]
                    cursor = connection.cursor()
                    # lista_json = []
                    # data = {}
                    sql = "select p.apellido1 , p.apellido2, p.nombres,de.descripcion, dep.nombre, p.telefonoextension ,p.emailinst,p.cedula,p.id " \
                          " from sga_persona p, auth_user usua, sagest_distributivopersona d, sagest_nivelocupacional nio, sagest_modalidadlaboral mod, sagest_regimenlaboral r, sagest_denominacionpuesto de, sagest_departamento dep " \
                          " where p.id=d.persona_id and d.status=true and d.estadopuesto_id=1 and r.id=d.regimenlaboral_id " \
                          " and usua.id=p.usuario_id and d.nivelocupacional_id=nio.id and d.modalidadlaboral_id=mod.id and de.id=d.denominacionpuesto_id and dep.id=d.unidadorganica_id order by p.apellido1,p.apellido2, p.nombres "
                    cursor.execute(sql)
                    results = cursor.fetchall()
                    row_num = 4
                    a = 0
                    for r in results:
                        i = 0
                        campo1 = i
                        campo2 = r[0]
                        campo3 = r[1]
                        campo4 = r[2]
                        campo5 = r[3]
                        campo6 = r[4]
                        campo7 = r[5]
                        campo8 = r[6]
                        campo9 = r[7]
                        # campo10 = r[8]
                        a += 1
                        ws.write(row_num, 0, a, font_style2)
                        # ws.write(row_num, 1, campo10, font_style2)
                        ws.write(row_num, 1, campo9, font_style2)
                        ws.write(row_num, 2, campo2 + ' ' + campo3 + ' ' + campo4, font_style2)
                        ws.write(row_num, 3, campo5, font_style2)
                        ws.write(row_num, 4, campo6, style1)
                        ws.write(row_num, 5, 'Cdla. Universitaria Km. 1 1/2 vía Km. 26', font_style2)
                        ws.write(row_num, 6, 'MILAGRO', font_style2)
                        ws.write(row_num, 7, '(04) 2 715081 - 2715079', font_style2)
                        ws.write(row_num, 8, campo7, font_style2)
                        ws.write(row_num, 9, campo8, font_style2)
                        row_num += 1
                    wb.save(response)
                    connection.close()
                    return response
                except Exception as ex:
                    pass

            # MRL
            if action == 'info_academica':
                try:
                    __author__ = 'Unemi'
                    style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
                    style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
                    style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
                    title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                    style1 = easyxf(num_format_str='D-MMM-YY')
                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('info_academica')
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=info_academica.csv'
                    columns = [
                        (u"nivelInstruccionId", 1000),
                        (u"numeroRegistroCertificado", 1000),
                        (u"institucionEducativa", 1000),
                        (u"anioEstudios", 1000),
                        (u"tipoPeriodoId", 1000),
                        (u"areaConocimiento", 1000),
                        (u"egresado", 1000),
                        (u"titulo", 1000),
                        (u"paisId", 1000),
                        (u"servidorId", 1000),
                    ]
                    row_num = 1
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], font_style)
                        ws.col(col_num).width = columns[col_num][1]
                    cursor = connection.cursor()
                    # lista_json = []
                    # data = {}
                    sql = "select t.codigo_tthh as nivelInstruccionId,ti.registro as numeroRegistroCertificado, " \
                          "(case ti.educacionsuperior when true then (select ies.nombre from sga_institucioneducacionsuperior ies where ies.id=ti.institucion_id) " \
                          " else (select co.nombre from sga_colegio co where co.id=ti.colegio_id) end) as institucionEducativa, " \
                          " ti.anios as anioEstudios, (case when ti.semestres>0 then 2 else 1 end) as tipoPeriodoId, " \
                          " COALESCE((select act.nombre from sga_areaconocimientotitulacion act where act.id=t.areaconocimiento_id),'') as areaConocimiento, " \
                          " (case COALESCE(CAST(to_char(ti.fechaegresado,'YYYY-MM-DD') as text),'') when '' then 0 else 1 end) as egresado, " \
                          " t.nombre as titulo, pa.codigo_tthh as paisId, p.cedula as servidorId from sga_titulacion ti, sga_persona p, sga_titulo t, sga_pais pa " \
                          " where ti.status=true and ti.persona_id=p.id and p.status=true and t.status=true and t.id=ti.titulo_id and ti.pais_id=pa.id and pa.status=true;"
                    cursor.execute(sql)
                    results = cursor.fetchall()
                    row_num = 2
                    a = 0
                    for r in results:
                        i = 0
                        campo1 = r[0]
                        campo2 = r[1]
                        campo3 = r[2]
                        campo4 = r[3]
                        campo5 = r[4]
                        campo6 = r[5]
                        campo7 = r[6]
                        campo8 = r[7]
                        campo9 = r[8]
                        campo10 = r[9]
                        a += 1
                        ws.write(row_num, 0, campo1, font_style2)
                        ws.write(row_num, 1, campo2, font_style2)
                        ws.write(row_num, 2, campo3, font_style2)
                        ws.write(row_num, 3, campo4, font_style2)
                        ws.write(row_num, 4, campo5, font_style2)
                        ws.write(row_num, 5, campo6, font_style2)
                        ws.write(row_num, 6, campo7, font_style2)
                        ws.write(row_num, 7, campo8, font_style2)
                        ws.write(row_num, 8, campo9, font_style2)
                        ws.write(row_num, 9, campo10, font_style2)
                        row_num += 1
                    wb.save(response)
                    connection.close()
                    return response
                except Exception as ex:
                    pass

            if action == 'info_capacitacion':
                try:
                    __author__ = 'Unemi'
                    style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
                    style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
                    style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
                    title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                    style1 = easyxf(num_format_str='D-MMM-YY')
                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('info_capacitacion')
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=info_capacitacion.csv'
                    columns = [
                        (u"evento", 1000),
                        (u"tipoEventoCapacitacion", 1000),
                        (u"auspiciante", 1000),
                        (u"duracion", 1000),
                        (u"tipoCertificado", 1000),
                        (u"certificadoPor", 1000),
                        (u"fechaInicio", 1000),
                        (u"fechaFin", 1000),
                        (u"pais", 1000),
                        (u"servidor", 1000),
                    ]
                    row_num = 1
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], font_style)
                        ws.col(col_num).width = columns[col_num][1]
                    cursor = connection.cursor()
                    # lista_json = []
                    # data = {}
                    sql = "select ca.nombre as evento, tc.codigo_tthh as tipoEventoCapacitacion , ca.auspiciante as auspiciante, ca.horas as duracion,  " \
                          " (case when COALESCE((select tp.nombre from sga_tipoparticipacion tp where ca.tipoparticipacion_id=tp.id and tp.status=true),'ASISTENCIA')='ASISTENCIA' then 2 else 1 end) as tipoCertificado, " \
                          " '' as certificadoPor, COALESCE(to_char(ca.fechainicio,'DD/MM/YYYY'),'') as fechaInicio, COALESCE(to_char(ca.fechafin,'DD/MM/YYYY'),'') as fechaFin,  " \
                          " (select pa.codigo_tthh from sga_pais pa where pa.id=ca.pais_id and pa.status=true) as pais, pe.cedula as	servidor " \
                          " from sga_capacitacion ca, sga_tipocurso tc , sga_persona pe where ca.status=true and ca.tipocurso_id=tc.id and tc.status=true " \
                          " and pe.id=ca.persona_id and pe.status=true;"
                    cursor.execute(sql)
                    results = cursor.fetchall()
                    row_num = 2
                    a = 0
                    for r in results:
                        i = 0
                        campo1 = r[0]
                        campo2 = r[1]
                        campo3 = r[2]
                        campo4 = r[3]
                        campo5 = r[4]
                        campo6 = r[5]
                        campo7 = r[6]
                        campo8 = r[7]
                        campo9 = r[8]
                        campo10 = r[9]
                        a += 1
                        ws.write(row_num, 0, campo1, font_style2)
                        ws.write(row_num, 1, campo2, font_style2)
                        ws.write(row_num, 2, campo3, font_style2)
                        ws.write(row_num, 3, campo4, font_style2)
                        ws.write(row_num, 4, campo5, font_style2)
                        ws.write(row_num, 5, campo6, font_style2)
                        ws.write(row_num, 6, campo7, font_style2)
                        ws.write(row_num, 7, campo8, font_style2)
                        ws.write(row_num, 8, campo9, font_style2)
                        ws.write(row_num, 9, campo10, font_style2)
                        row_num += 1
                    wb.save(response)
                    connection.close()
                    return response
                except Exception as ex:
                    pass

            if action == 'info_carga_familiar':
                try:
                    __author__ = 'Unemi'
                    style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
                    style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
                    style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
                    title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                    style1 = easyxf(num_format_str='D-MMM-YY')
                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('info_carga_familiar')
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=info_carga_familiar.csv'
                    columns = [
                        (u"tipoDocumentoId", 1000),
                        (u"numeroDocumento", 1000),
                        (u"nombre", 1000),
                        (u"apellido", 1000),
                        (u"fechaNacimiento", 1000),
                        (u"nivelInstruccionId", 1000),
                        (u"servidorId", 1000),
                    ]
                    row_num = 1
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], font_style)
                        ws.col(col_num).width = columns[col_num][1]
                    cursor = connection.cursor()
                    # lista_json = []
                    # data = {}
                    sql = "select 1 as tipoDocumentoId, pdf.identificacion as	numeroDocumento, pdf.nombre as nombre,'' as	apellido, " \
                          " COALESCE(to_char(pdf.nacimiento,'DD/MM/YYYY'),'') as	fechaNacimiento,nt.codigo_tthh as	nivelInstruccionId,pe.cedula as servidorId " \
                          " from sga_personadatosfamiliares pdf, sga_persona pe, sga_niveltitulacion nt " \
                          " where pdf.status=true and pdf.persona_id=pe.id and pe.status=true " \
                          " and pe.id in (select ad.persona_id from sga_administrativo ad where ad.status=true) " \
                          " and pe.id in (select pr.persona_id from sga_profesor pr where pr.status=true) " \
                          " and nt.id=pdf.niveltitulacion_id and nt.status=true;"
                    cursor.execute(sql)
                    results = cursor.fetchall()
                    row_num = 2
                    a = 0
                    for r in results:
                        i = 0
                        campo1 = r[0]
                        campo2 = r[1]
                        campo3 = r[2]
                        campo4 = r[3]
                        campo5 = r[4]
                        campo6 = r[5]
                        campo7 = r[6]
                        a += 1
                        ws.write(row_num, 0, campo1, font_style2)
                        ws.write(row_num, 1, campo2, font_style2)
                        ws.write(row_num, 2, campo3, font_style2)
                        ws.write(row_num, 3, campo4, font_style2)
                        ws.write(row_num, 4, campo5, font_style2)
                        ws.write(row_num, 5, campo6, font_style2)
                        ws.write(row_num, 6, campo7, font_style2)
                        row_num += 1
                    wb.save(response)
                    connection.close()
                    return response
                except Exception as ex:
                    pass

            if action == 'info_carga_mas_info_personal':
                try:
                    __author__ = 'Unemi'
                    style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
                    style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
                    style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
                    title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                    style1 = easyxf(num_format_str='D-MMM-YY')
                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('info_carga_mas_info_personal')
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=info_carga_mas_info_personal.csv'
                    columns = [
                        (u"numeroDocumento", 1000),
                        (u"apellido", 1000),
                        (u"nombre", 1000),
                        (u"tipoDocumento", 1000),
                        (u"servidorPasanteConvenio", 1000),
                        (u"numeroLibretaMilitar", 1000),
                        (u"nacionalidadId", 1000),
                        (u"aniosResidencia", 1000),
                        (u"fechaNacimiento", 1000),
                        (u"sexoId", 1000),
                        (u"tipoSangreId", 1000),
                        (u"estadoCivilId", 1000),
                        (u"discapacidad", 1000),
                        (u"numeroCarnetConadis", 1000),
                        (u"tipoDiscapacidadId", 1000),
                        (u"servidorCarrera", 1000),
                        (u"numeroRegistroCertificado", 1000),
                        (u"identificacionEtnicaId", 1000),
                        (u"nacionalidadIndigenaId", 1000),
                        (u"direccionCallePrincipal", 1000),
                        (u"direccionNumero", 1000),
                        (u"direccionCalleSecundaria", 1000),
                        (u"direccionReferencia", 1000),
                        (u"telefonoDomicilio", 1000),
                        (u"telefonoCelular", 1000),
                        (u"telefonoTrabajo", 1000),
                        (u"telefonoExtension", 1000),
                        (u"correoElectronico", 1000),
                        (u"correoElectronicoTmp", 1000),
                        (u"direccionProvinciaId", 1000),
                        (u"direccionCantonId", 1000),
                        (u"direccionParroquiaId", 1000),
                        (u"contactoApellido", 1000),
                        (u"contactoNombre", 1000),
                        (u"contactoTelefono", 1000),
                        (u"contactoTelefonoCelular", 1000),
                        (u"numeroNotariaBienes", 1000),
                        (u"lugarNotariaBienesId", 1000),
                        (u"fechaNotariaBienes", 1000),
                    ]
                    row_num = 1
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], font_style)
                        ws.col(col_num).width = columns[col_num][1]
                    cursor = connection.cursor()
                    # lista_json = []
                    # data = {}
                    sql = "select p.cedula as numeroDocumento, p.apellido1||' '||p.apellido2 as apellido, p.nombres as nombre, 1 as tipoDocumento, 0 as servidorPasanteConvenio , " \
                          " p.libretamilitar as numeroLibretaMilitar, (select pa.codigonacionalidad from sga_pais pa where pa.id=p.pais_id) as nacionalidadId, " \
                          " p.anioresidencia as aniosResidencia, to_char(p.nacimiento,'DD/MM/YYYY') as fechaNacimiento, (case p.sexo_id when 1 then 0 else 1 end) as sexoId, " \
                          " (select ts.codigo_tthh from sga_tiposangre ts where ts.id=p.sangre_id) as tipoSangreId, " \
                          " (select estado.codigo_tthh  from med_personaextension medpersona, sga_personaestadocivil estado where medpersona.estadocivil_id=estado.id and medpersona.persona_id=p.id and medpersona.status=True) as estadoCivilId, " \
                          " (select (case perfil.tienediscapacidad when True then 1 else 0 end) from sga_perfilinscripcion perfil where perfil.persona_id=p.id limit 1) as discapacidad, " \
                          " (select perfil.carnetdiscapacidad from sga_perfilinscripcion perfil where perfil.persona_id=p.id limit 1) as numeroCarnetConadis, " \
                          " (select di.codigo_tthh from sga_perfilinscripcion perfil,sga_discapacidad di where perfil.tipodiscapacidad_id=di.id and perfil.persona_id=p.id limit 1) as discapacidad, " \
                          " (case p.servidorcarrera when True then 1 else 0 end) as servidorCarrera, p.regitrocertificacion as numeroRegistroCertificado, " \
                          " (select raz.codigo_tthh from sga_perfilinscripcion perfil,sga_raza raz where perfil.raza_id=raz.id and perfil.persona_id=p.id limit 1) as identificacionEtnicaId, " \
                          " (select ni.codigo_tthh from sga_perfilinscripcion perfil,sga_nacionalidadindigena ni where perfil.nacionalidadindigena_id=ni.id and perfil.persona_id=p.id limit 1) as nacionalidadIndigenaId, " \
                          " p.direccion as direccionCallePrincipal, p.num_direccion as direccionNumero, p.direccion2 as direccionCalleSecundaria,p.referencia as direccionReferencia, " \
                          " p.telefono_conv as telefonoDomicilio, p.telefono as telefonoCelular, '' as telefonoTrabajo, p.telefonoextension as telefonoExtension, " \
                          " p.emailinst as correoElectronico, p.email as correoElectronicoTmp, " \
                          " COALESCE((select pr.codigo_tthh from sga_provincia pr where pr.id=p.provincia_id),0) as direccionProvinciaId, " \
                          " COALESCE((select can.codigo_tthh from sga_canton can where can.id=p.canton_id),0) as direccionCantonId, " \
                          " COALESCE((select par.codigo_tthh from sga_parroquia par where par.id=p.parroquia_id),0) as direccionParroquiaId, " \
                          " COALESCE((select pex.contactoemergencia from med_personaextension pex where pex.persona_id=p.id),'') as contactoApellido,'' as contactoNombre, " \
                          " COALESCE((select pex.telefonoemergencia from med_personaextension pex where pex.persona_id=p.id),'') as contactoTelefono, " \
                          " '' as contactoTelefonoCelular, COALESCE((select db.numero from sga_declaracionbienes db where db.persona_id=p.id and db.status=true  limit 1),'') as numeroNotariaBienes, " \
                          " '' as lugarNotariaBienesId, COALESCE(to_char((select db.fecha from sga_declaracionbienes db where db.persona_id=p.id and db.status=true  limit 1),'DD/MM/YYYY'),'') as fechaNotariaBienes " \
                          " from sga_persona p, sagest_distributivopersona d " \
                          " where p.id=d.persona_id and d.status=true and d.estadopuesto_id=1 and p.status=true;"
                    cursor.execute(sql)
                    results = cursor.fetchall()
                    row_num = 2
                    a = 0
                    for r in results:
                        i = 0
                        campo1 = r[0]
                        campo2 = r[1]
                        campo3 = r[2]
                        campo4 = r[3]
                        campo5 = r[4]
                        campo6 = r[5]
                        campo7 = r[6]
                        campo8 = r[7]
                        campo9 = r[8]
                        campo10 = r[9]
                        campo11 = r[10]
                        campo12 = r[11]
                        campo13 = r[12]
                        campo14 = r[13]
                        campo15 = r[14]
                        campo16 = r[15]
                        campo17 = r[16]
                        campo18 = r[17]
                        campo19 = r[18]
                        campo20 = r[19]
                        campo21 = r[20]
                        campo22 = r[21]
                        campo23 = r[22]
                        campo24 = r[23]
                        campo25 = r[24]
                        campo26 = r[25]
                        campo27 = r[26]
                        campo28 = r[27]
                        campo29 = r[28]
                        campo30 = r[29]
                        campo31 = r[30]
                        campo32 = r[31]
                        campo33 = r[32]
                        campo34 = r[33]
                        campo35 = r[34]
                        campo36 = r[35]
                        campo37 = r[36]
                        campo38 = r[37]
                        campo39 = r[38]
                        a += 1
                        ws.write(row_num, 0, campo1, font_style2)
                        ws.write(row_num, 1, campo2, font_style2)
                        ws.write(row_num, 2, campo3, font_style2)
                        ws.write(row_num, 3, campo4, font_style2)
                        ws.write(row_num, 4, campo5, font_style2)
                        ws.write(row_num, 5, campo6, font_style2)
                        ws.write(row_num, 6, campo7, font_style2)
                        ws.write(row_num, 7, campo8, font_style2)
                        ws.write(row_num, 8, campo9, font_style2)
                        ws.write(row_num, 9, campo10, font_style2)
                        ws.write(row_num, 10, campo11, font_style2)
                        ws.write(row_num, 11, campo12, font_style2)
                        ws.write(row_num, 12, campo13, font_style2)
                        ws.write(row_num, 13, campo14, font_style2)
                        ws.write(row_num, 14, campo15, font_style2)
                        ws.write(row_num, 15, campo16, font_style2)
                        ws.write(row_num, 16, campo17, font_style2)
                        ws.write(row_num, 17, campo18, font_style2)
                        ws.write(row_num, 18, campo19, font_style2)
                        ws.write(row_num, 19, campo20, font_style2)
                        ws.write(row_num, 20, campo21, font_style2)
                        ws.write(row_num, 21, campo22, font_style2)
                        ws.write(row_num, 22, campo23, font_style2)
                        ws.write(row_num, 23, campo24, font_style2)
                        ws.write(row_num, 24, campo25, font_style2)
                        ws.write(row_num, 25, campo26, font_style2)
                        ws.write(row_num, 26, campo27, font_style2)
                        ws.write(row_num, 27, campo28, font_style2)
                        ws.write(row_num, 28, campo29, font_style2)
                        ws.write(row_num, 29, campo30, font_style2)
                        ws.write(row_num, 30, campo31, font_style2)
                        ws.write(row_num, 31, campo32, font_style2)
                        ws.write(row_num, 32, campo33, font_style2)
                        ws.write(row_num, 33, campo34, font_style2)
                        ws.write(row_num, 34, campo35, font_style2)
                        ws.write(row_num, 35, campo36, font_style2)
                        ws.write(row_num, 36, campo37, font_style2)
                        ws.write(row_num, 37, campo38, font_style2)
                        ws.write(row_num, 38, campo39, font_style2)
                        row_num += 1
                    wb.save(response)
                    connection.close()
                    return response
                except Exception as ex:
                    pass

            if action == 'info_conyuge':
                try:
                    __author__ = 'Unemi'
                    style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
                    style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
                    style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
                    title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                    style1 = easyxf(num_format_str='D-MMM-YY')
                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('info_conyuge')
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=info_conyuge.csv'
                    columns = [
                        (u"tipoDocumentoId", 1000),
                        (u"numeroDocumento", 1000),
                        (u"nombre", 1000),
                        (u"apellido", 1000),
                        (u"tipoRelacionId", 1000),
                        (u"servidorId", 1000),
                    ]
                    row_num = 1
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], font_style)
                        ws.col(col_num).width = columns[col_num][1]
                    cursor = connection.cursor()
                    # lista_json = []
                    # data = {}
                    sql = "select (case (COALESCE(pdf.identificacion,'')) when '' then '' else '1' end) as tipoDocumentoId,  " \
                          " pdf.identificacion as numeroDocumento, pdf.nombre as nombre, '' as apellido, " \
                          " 2 as tipoRelacionId, p.cedula as servidorId " \
                          " from sga_persona p, sagest_distributivopersona d , sga_personadatosfamiliares pdf " \
                          " where p.id=d.persona_id and d.status=true and d.estadopuesto_id=1 and p.status=true " \
                          " and pdf.persona_id=p.id and pdf.status=true and pdf.parentesco_id=13;"
                    cursor.execute(sql)
                    results = cursor.fetchall()
                    row_num = 2
                    a = 0
                    for r in results:
                        i = 0
                        campo1 = r[0]
                        campo2 = r[1]
                        campo3 = r[2]
                        campo4 = r[3]
                        campo5 = r[4]
                        campo6 = r[5]
                        ws.write(row_num, 0, campo1, font_style2)
                        ws.write(row_num, 1, campo2, font_style2)
                        ws.write(row_num, 2, campo3, font_style2)
                        ws.write(row_num, 3, campo4, font_style2)
                        ws.write(row_num, 4, campo5, font_style2)
                        ws.write(row_num, 5, campo6, font_style2)
                        row_num += 1
                    wb.save(response)
                    connection.close()
                    return response
                except Exception as ex:
                    pass

            if action == 'info_cta_bancarias':
                try:
                    __author__ = 'Unemi'
                    style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
                    style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
                    style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
                    title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                    style1 = easyxf(num_format_str='D-MMM-YY')
                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('info_cta_bancarias')
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=info_cta_bancarias.csv'
                    columns = [
                        (u"numeroDocumento", 1000),
                        (u"institucionFinancieraBancariaId", 1000),
                        (u"tipoCuentaBancariaId", 1000),
                        (u"numeroCuentaBancaria", 1000),
                    ]
                    row_num = 1
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], font_style)
                        ws.col(col_num).width = columns[col_num][1]
                    cursor = connection.cursor()
                    # lista_json = []
                    # data = {}
                    sql = "select p.cedula as numeroDocumento, b.codigo_tthh as institucionFinancieraBancariaId,  " \
                          " tcb.codigo as tipoCuentaBancariaId, cbp.numero as	numeroCuentaBancaria " \
                          " from sga_persona p, sagest_distributivopersona d , sga_cuentabancariapersona cbp, sagest_banco b, " \
                          " sagest_tipocuentabanco tcb " \
                          " where p.id=d.persona_id and d.status=true and d.estadopuesto_id=1 and p.status=true " \
                          " and cbp.persona_id=p.id and cbp.status=true and b.id=cbp.banco_id and b.status=true and " \
                          " cbp.tipocuentabanco_id=tcb.id and tcb.status=true; "
                    cursor.execute(sql)
                    results = cursor.fetchall()
                    row_num = 2
                    a = 0
                    for r in results:
                        i = 0
                        campo1 = r[0]
                        campo2 = r[1]
                        campo3 = r[2]
                        campo4 = r[3]
                        ws.write(row_num, 0, campo1, font_style2)
                        ws.write(row_num, 1, campo2, font_style2)
                        ws.write(row_num, 2, campo3, font_style2)
                        ws.write(row_num, 3, campo4, font_style2)
                        row_num += 1
                    wb.save(response)
                    connection.close()
                    return response
                except Exception as ex:
                    pass

            if action == 'info_evaluacion_desempenio':
                try:
                    __author__ = 'Unemi'
                    style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
                    style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
                    style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
                    title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                    style1 = easyxf(num_format_str='D-MMM-YY')
                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('info_evaluacion_desempenio')
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=info_evaluacion_desempenio.csv'
                    columns = [
                        (u"servidorPuestoId", 1000),
                        (u"fechaEvaluacionInicio", 1000),
                        (u"fechaEvaluacionFin", 1000),
                        (u"puntaje", 1000),
                        (u"calificacionId", 1000),
                    ]
                    row_num = 1
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], font_style)
                        ws.col(col_num).width = columns[col_num][1]
                    cursor = connection.cursor()
                    # lista_json = []
                    # data = {}
                    sql = "select p.cedula as servidorPuestoId, pe.inicioeval as fechaEvaluacionInicio, pe.fineval as fechaEvaluacionFin, " \
                          " pe.totaleva as	puntaje, (case when pe.totaleva >= 90.5 then 1 when pe.totaleva >= 80.5 then 3 when pe.totaleva >= 70.5 then 2 " \
                          " when pe.totaleva >= 60.5 then 8 else 6 end) as calificacionId " \
                          " from sagest_podevaluaciondet pe, sga_persona p " \
                          " where pe.status=true and p.id=pe.evaluado_id;"
                    cursor.execute(sql)
                    results = cursor.fetchall()
                    row_num = 2
                    a = 0
                    for r in results:
                        i = 0
                        campo1 = r[0]
                        campo2 = r[1]
                        campo3 = r[2]
                        campo4 = r[3]
                        campo5 = r[4]
                        ws.write(row_num, 0, campo1, font_style2)
                        ws.write(row_num, 1, campo2, font_style2)
                        ws.write(row_num, 2, campo3, font_style2)
                        ws.write(row_num, 3, campo4, font_style2)
                        ws.write(row_num, 4, campo5, font_style2)
                        row_num += 1
                    wb.save(response)
                    connection.close()
                    return response
                except Exception as ex:
                    pass

            if action == 'info_trayectoria_laboral':
                try:
                    __author__ = 'Unemi'
                    style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
                    style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
                    style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
                    title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
                    style1 = easyxf(num_format_str='D-MMM-YY')
                    font_style = XFStyle()
                    font_style.font.bold = True
                    font_style2 = XFStyle()
                    font_style2.font.bold = False
                    wb = Workbook(encoding='utf-8')
                    ws = wb.add_sheet('info_trayectoria_laboral')
                    response = HttpResponse(content_type="application/ms-excel")
                    response['Content-Disposition'] = 'attachment; filename=info_trayectoria_laboral.csv'
                    columns = [
                        (u"tipoInstitucion", 1000),
                        (u"institucion", 1000),
                        (u"unidadAdministrativa", 1000),
                        (u"denominacionPuesto", 1000),
                        (u"fechaIngreso", 1000),
                        (u"fechaSalida", 1000),
                        (u"motivoIngreso", 1000),
                        (u"motivoSalida", 1000),
                        (u"servidor", 1000),
                    ]
                    row_num = 1
                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], font_style)
                        ws.col(col_num).width = columns[col_num][1]
                    cursor = connection.cursor()
                    # lista_json = []
                    # data = {}
                    sql = "select e.tipoinstitucion as tipoInstitucion, e.institucion as institucion, e.departamento as unidadAdministrativa,  " \
                          " e.cargo as denominacionPuesto, COALESCE(to_char(e.fechainicio,'DD/MM/YYYY'),'') as fechaIngreso,  " \
                          " COALESCE(to_char(e.fechafin,'DD/MM/YYYY'),'') as fechaSalida, '' as motivoIngreso, COALESCE((select ms.codigo_tthh from sagest_motivosalida ms where ms.id=e.motivosalida_id),0) as motivoSalida, p.cedula as servidor " \
                          " from sagest_experiencialaboral e, sga_persona p where e.status=true and e.persona_id=p.id;"
                    cursor.execute(sql)
                    results = cursor.fetchall()
                    row_num = 2
                    a = 0
                    for r in results:
                        i = 0
                        campo1 = r[0]
                        campo2 = r[1]
                        campo3 = r[2]
                        campo4 = r[3]
                        campo5 = r[4]
                        campo6 = r[5]
                        campo7 = r[6]
                        campo8 = r[7]
                        campo9 = r[8]
                        ws.write(row_num, 0, campo1, font_style2)
                        ws.write(row_num, 1, campo2, font_style2)
                        ws.write(row_num, 2, campo3, font_style2)
                        ws.write(row_num, 3, campo4, font_style2)
                        ws.write(row_num, 4, campo5, font_style2)
                        ws.write(row_num, 5, campo6, font_style2)
                        ws.write(row_num, 6, campo7, font_style2)
                        ws.write(row_num, 7, campo8, font_style2)
                        ws.write(row_num, 8, campo9, font_style2)
                        row_num += 1
                    wb.save(response)
                    connection.close()
                    return response
                except Exception as ex:
                    pass

            return HttpResponseRedirect(request.path)
        else:
            data['title'] = u'Distributivo de la institucion.'
            search = None
            ids = None
            if 's' in request.GET:
                search = request.GET['s'].strip()
                ss = search.split(' ')
                if len(ss) == 1:
                    plantillas = DistributivoPersona.objects.filter(Q(persona__nombres__icontains=search) |
                                                                    Q(persona__apellido1__icontains=search) |
                                                                    Q(persona__apellido2__icontains=search) |
                                                                    Q(persona__cedula__icontains=search) |
                                                                    Q(persona__pasaporte__icontains=search)).distinct()
                else:
                    plantillas = DistributivoPersona.objects.filter(Q(persona__apellido1__icontains=ss[0]) & Q(persona__apellido2__icontains=ss[1])).distinct()
            elif 'id' in request.GET:
                ids = request.GET['id']
                plantillas = DistributivoPersona.objects.filter(id=ids)
            else:
                plantillas = DistributivoPersona.objects.all()
            paging = MiPaginador(plantillas, 20)
            p = 1
            try:
                paginasesion = 1
                if 'paginador' in request.session:
                    paginasesion = int(request.session['paginador'])
                if 'page' in request.GET:
                    p = int(request.GET['page'])
                else:
                    p = paginasesion
                try:
                    page = paging.page(p)
                except:
                    p = 1
                page = paging.page(p)
            except:
                page = paging.page(p)
            request.session['paginador'] = p
            data['paging'] = paging
            data['rangospaging'] = paging.rangos_paginado(p)
            data['page'] = page
            data['search'] = search if search else ""
            data['ids'] = ids if ids else ""
            data['plantillas'] = page.object_list
            data['email_domain'] = EMAIL_DOMAIN
            return render(request, 'th_plantilla/view.html', data)
