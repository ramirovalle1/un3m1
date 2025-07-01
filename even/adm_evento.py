# -*- coding: UTF-8 -*-
import io
import random
import sys

import xlsxwriter
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Q
import xlwt
from django.forms import model_to_dict
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.template import Context
from django.template.loader import get_template
from django.views.decorators.csrf import csrf_exempt
from xlwt import *
from django.shortcuts import render
from decorators import secure_module, last_access
from settings import EMAIL_DOMAIN

from sga.commonviews import adduserdata
from sga.forms import CargarMuestraForm

from sga.funciones import MiPaginador, log, generar_nombre
from sga.models import Persona
from .forms import MantenimientoForm, PeriodoEventoForm
from .models import *
from django.db.models import Value, Count, Sum, F, FloatField
from django.db.models import Count, Case, When, CharField, F
from django.db.models.functions import Coalesce

# @csrf_exempt
@login_required(redirect_field_name='ret', login_url='/loginsagest')
@secure_module
@last_access
@transaction.atomic()
def view(request):
    data = {}
    adduserdata(request, data)
    usuario = request.user
    if request.method == 'POST':
        action = request.POST['action']

        if action == 'mostrarperiodo':
            try:
                evento = PeriodoEvento.objects.get(pk=request.POST['id'])
                evento.publicar = True if request.POST['val'] == 'y' else False
                evento.save(request)
                log(u'Publica periodo en  : %s (%s)' % (evento, evento.publicar),
                    request, "edit")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad"})

        if action == 'addevento':
            try:
                with transaction.atomic():
                    form = MantenimientoForm(request.POST)
                    if form.is_valid():
                        evento = Evento(nombre=form.cleaned_data['nombre'].upper())
                        evento.save(request)
                        log(u'Adicionó Evento: %s' % evento, request, "add")
                        # return JsonResponse({"result": False,'to':'/nuevaurl'}, safe=False) SI DESEAS REDIRECCIONAR ADICIONARLE TO A LA RESPUESTA
                        return JsonResponse({"result": False}, safe=False)
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({"result": True, "mensaje": "Complete los datos requeridos."}, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        if action == 'addtipo':
            try:
                with transaction.atomic():
                    form = MantenimientoForm(request.POST)
                    if form.is_valid():
                        tipo = TipoEvento(nombre=form.cleaned_data['nombre'].upper())
                        tipo.save(request)
                        log(u'Adicionó Tipo de Evento: %s' % tipo, request, "add")
                        # return JsonResponse({"result": False,'to':'/nuevaurl'}, safe=False) SI DESEAS REDIRECCIONAR ADICIONARLE TO A LA RESPUESTA
                        return JsonResponse({"result": False}, safe=False)
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({"result": True, "mensaje": "Complete los datos requeridos."}, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        if action == 'addperiodo':
            try:
                with transaction.atomic():
                    form = PeriodoEventoForm(request.POST, request.FILES)
                    if form.is_valid():
                        periodo = PeriodoEvento(evento=form.cleaned_data['evento'],
                                                tipo=form.cleaned_data['tipo'],
                                                descripcionbreve=form.cleaned_data['descripcionbreve'].upper(),
                                                fechainicio=form.cleaned_data['fechainicio'],
                                                fechainicioinscripcion=form.cleaned_data['fechainicioinscripcion'],
                                                fechainicioconfirmar=form.cleaned_data['fechainicioconfirmar'],
                                                fechafin=form.cleaned_data['fechafin'],
                                                fechafininscripcion=form.cleaned_data['fechafininscripcion'],
                                                fechafinconfirmar=form.cleaned_data['fechafinconfirmar'],
                                                horainicio=form.cleaned_data['horainicio'],
                                                horafin=form.cleaned_data['horafin'],
                                                cuerpo=form.cleaned_data['cuerpo'],
                                                iframemapa=form.cleaned_data['iframemapa'],
                                                publicar=form.cleaned_data['publicar'],
                                                permiteregistro=form.cleaned_data['permiteregistro'],
                                                todos=form.cleaned_data['todos'],
                                                aplicaperiodo=form.cleaned_data['aplicaperiodo'],
                                                validacupo=form.cleaned_data['validacupo'],
                                                link=form.cleaned_data['link'],
                                                tipoperfil=form.cleaned_data['tipoperfil'])
                        if 'imagen' in request.FILES:
                            newfile = request.FILES['imagen']
                            if newfile:
                                if newfile.size > 4194304:
                                    return JsonResponse({"result": "bad", "mensaje": u"Error, archivo mayor a 3 Mb."})
                                else:
                                    newfilesd = newfile._name
                                    ext = newfilesd[newfilesd.rfind("."):]
                                    if ext in ['.jpg', '.jpeg', '.png']:
                                        newfile._name = generar_nombre("imagen_evento", newfile._name)
                                        periodo.imagen = newfile

                                    else:
                                        return JsonResponse({"result": "bad",
                                                             "mensaje": u"Error, Solo archivo con extensión. pdf, jpg, jpeg."})
                        if 'portada' in request.FILES:
                            newfile2 = request.FILES['portada']
                            if newfile2:
                                if newfile2.size > 4194304:
                                    return JsonResponse({"result": "bad", "mensaje": u"Error, archivo mayor a 3 Mb."})
                                else:
                                    newfilesd = newfile2._name
                                    ext = newfilesd[newfilesd.rfind("."):]
                                    if ext in ['.jpg', '.jpeg', '.png']:
                                        newfile2._name = generar_nombre("imagen_evento", newfile2._name)
                                        periodo.portada = newfile2

                                    else:
                                        return JsonResponse({"result": "bad",
                                                             "mensaje": u"Error, Solo archivo con extensión. pdf, jpg, jpeg."})
                        periodo.save(request)

                        if form.cleaned_data['aplicaperiodo']:
                            periodo.periodo = form.cleaned_data['periodo']
                            periodo.save(request)

                        if form.cleaned_data['validacupo']:
                            periodo.cupo = form.cleaned_data['cupo']
                            periodo.save(request)

                        if not form.cleaned_data['todos']:
                            cantones = request.POST.getlist('cantones')
                            for canton in cantones:
                                detalle = DetallePeriodoEvento(periodo=periodo,
                                                               canton_id =int(canton) )
                                detalle.save(request)

                        log(u'Adicionó Tipo de Evento: %s' % periodo, request, "add")
                        return JsonResponse({"result": "ok"})
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({"result": True, "mensaje": "Complete los datos requeridos."}, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        if action == 'editevento':
            try:
                with transaction.atomic():
                    form = MantenimientoForm(request.POST)
                    if form.is_valid():
                        evento = Evento.objects.get(id=request.POST['id'])
                        evento.nombre =form.cleaned_data['nombre'].upper()
                        evento.save(request)
                        log(u'Editó Evento: %s' % evento, request, "edit")
                        # return JsonResponse({"result": False,'to':'/nuevaurl'}, safe=False) SI DESEAS REDIRECCIONAR ADICIONARLE TO A LA RESPUESTA
                        return JsonResponse({"result": False}, safe=False)
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({"result": True, "mensaje": "Complete los datos requeridos."}, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        if action == 'edittipo':
            try:
                with transaction.atomic():
                    form = MantenimientoForm(request.POST)
                    if form.is_valid():
                        tipo = TipoEvento.objects.get(id=request.POST['id'])
                        tipo.nombre =form.cleaned_data['nombre'].upper()
                        tipo.save(request)
                        log(u'Editó Tipo de Evento: %s' % tipo, request, "edit")
                        # return JsonResponse({"result": False,'to':'/nuevaurl'}, safe=False) SI DESEAS REDIRECCIONAR ADICIONARLE TO A LA RESPUESTA
                        return JsonResponse({"result": False}, safe=False)
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({"result": True, "mensaje": "Complete los datos requeridos."}, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        if action == 'editperiodo':
            try:
                with transaction.atomic():
                    form = PeriodoEventoForm(request.POST)
                    if form.is_valid():
                        periodo = PeriodoEvento.objects.get(id=request.POST['id'])

                        periodo.periodo = form.cleaned_data['periodo']
                        periodo.evento = form.cleaned_data['evento']
                        periodo.tipo = form.cleaned_data['tipo']
                        periodo.fechainicio = form.cleaned_data['fechainicio']
                        periodo.fechainicioinscripcion = form.cleaned_data['fechainicioinscripcion']
                        periodo.fechainicioconfirmar = form.cleaned_data['fechainicioconfirmar']
                        periodo.fechafin = form.cleaned_data['fechafin']
                        periodo.fechafininscripcion = form.cleaned_data['fechafininscripcion']
                        periodo.fechafinconfirmar = form.cleaned_data['fechafinconfirmar']
                        periodo.horainicio = form.cleaned_data['horainicio']
                        periodo.horafin = form.cleaned_data['horafin']
                        periodo.cuerpo = form.cleaned_data['cuerpo']
                        periodo.iframemapa = form.cleaned_data['iframemapa']
                        periodo.descripcionbreve = form.cleaned_data['descripcionbreve'].upper()
                        periodo.tipoperfil = form.cleaned_data['tipoperfil']
                        periodo.publicar = form.cleaned_data['publicar']
                        periodo.validacupo = form.cleaned_data['validacupo']
                        periodo.todos = form.cleaned_data['todos']
                        periodo.permiteregistro = form.cleaned_data['permiteregistro']
                        periodo.link = form.cleaned_data['link']

                        if 'imagen' in request.FILES:
                            newfile2 = request.FILES['imagen']
                            if newfile2:
                                if newfile2.size > 4194304:
                                    return JsonResponse({"result": "bad", "mensaje": u"Error, archivo mayor a 3 Mb."})
                                else:
                                    newfilesd = newfile2._name
                                    ext = newfilesd[newfilesd.rfind("."):]
                                    if ext in ['.jpg', '.jpeg', '.png']:
                                        newfile2._name = generar_nombre("imagen_evento", newfile2._name)
                                        periodo.imagen = newfile2

                                    else:
                                        return JsonResponse({"result": "bad",
                                                             "mensaje": u"Error, Solo archivo con extensión. jpg, jpeg."})

                        if 'portada' in request.FILES:
                            newfile2 = request.FILES['portada']
                            if newfile2:
                                if newfile2.size > 4194304:
                                    return JsonResponse({"result": "bad", "mensaje": u"Error, archivo mayor a 3 Mb."})
                                else:
                                    newfilesd = newfile2._name
                                    ext = newfilesd[newfilesd.rfind("."):]
                                    if ext in ['.jpg', '.jpeg', '.png']:
                                        newfile2._name = generar_nombre("imagen_evento", newfile2._name)
                                        periodo.portada = newfile2

                                    else:
                                        return JsonResponse({"result": "bad",
                                                             "mensaje": u"Error, Solo archivo con extensión. jpg, jpeg."})

                        periodo.save(request)

                        if form.cleaned_data['aplicaperiodo']:
                            periodo.periodo = form.cleaned_data['periodo']

                        if form.cleaned_data['validacupo']:
                            periodo.cupo = form.cleaned_data['cupo']
                        else:
                            periodo.cupo=0
                        periodo.save()
                        if not form.cleaned_data['todos']:
                            cantones = request.POST.getlist('cantones')
                            for canton in cantones:
                                if not DetallePeriodoEvento.objects.filter(periodo=periodo,canton_id=int(canton)).exists():
                                    detalle = DetallePeriodoEvento(periodo=periodo,
                                                                   canton_id =int(canton))
                                    detalle.save(request)
                        log(u'Editó Evento configurado: %s' % periodo, request, "edit")
                        return JsonResponse({"result": 'ok'})
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({"result": True, "mensaje": "Complete los datos requeridos."}, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        if action == 'deltipo':
            try:
                tipo = TipoEvento.objects.get(pk=request.POST['id'])
                tipo.status = False
                tipo.save(request)
                log(u'Elimino tipo de evento: %s' % tipo, request, "del")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        if action == 'delevento':
            try:
                evento = Evento.objects.get(pk=request.POST['id'])
                evento.status = False
                evento.save(request)
                log(u'Elimino evento: %s' % evento, request, "del")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        if action == 'delperiodo':
            try:
                evento = PeriodoEvento.objects.get(pk=request.POST['id'])
                evento.status = False
                evento.save(request)
                log(u'Eliminó periodo: %s' % evento, request, "del")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        elif action == 'cargarmuestra':
            try:
                with transaction.atomic():
                    import openpyxl
                    if not 'archivo' in request.FILES:
                        return JsonResponse({"result": True, "mensaje": "Debe cargar un archivo con muestra."}, safe=False)
                    excel = request.FILES['archivo']
                    id = request.POST['id']
                    periodo = PeriodoEvento.objects.get(pk=id)
                    wb = openpyxl.load_workbook(excel)
                    worksheet = wb.worksheets[0]
                    linea, excluido, cargados = 1, 0, 0
                    lista_excluidos = []
                    for row in worksheet.iter_rows():
                        currentValues = [str(cell.value) for cell in row]
                        if linea > 1:
                            if not currentValues[0] == 'None':
                                if Persona.objects.filter(status=True, cedula=currentValues[0]).exists():
                                    pers = Persona.objects.filter(status=True, cedula=currentValues[0]).first()
                                    if not RegistroEvento.objects.filter(status=True, participante=pers, periodo=periodo).exists():
                                        notimuestra = RegistroEvento(participante=pers, periodo=periodo)
                                        notimuestra.save(request)
                                        cargados += 1
                                    else:
                                        lista_excluidos.append(currentValues[0])
                                        excluido += 1
                                else:
                                    lista_excluidos.append(currentValues[0])
                                    excluido += 1
                            else:
                                excluido += 1
                        linea += 1
                    messages.success(request, 'Registro Guardado, se cargaron un total {} y se excluyeron un total de {}'.format(cargados, excluido))
                    return JsonResponse({"result": False}, safe=False)
            except Exception as ex:
                print(ex)
                print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        return JsonResponse({"result": "bad", "mensaje": u"Solicitud Incorrecta."})
    else:
        if 'action' in request.GET:
            data['action'] = action = request.GET['action']


            if action == 'addevento':
                try:
                    data['form2'] = MantenimientoForm()
                    template = get_template("adm_evento/modal/addevento.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            if action == 'addtipo':
                try:
                    data['form2'] = MantenimientoForm()
                    template = get_template("adm_evento/modal/addevento.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            if action == 'addperiodo':
                try:
                    data['title'] = u'Adicionar Evento Configurado'
                    data['form'] = PeriodoEventoForm()
                    return render(request, "adm_evento/modal/addperiodo.html", data)
                except Exception as ex:
                    pass

            if action == 'editevento':
                try:
                    data['id'] = request.GET['id']
                    data['filtro'] = filtro = Evento.objects.get(pk=request.GET['id'])
                    data['form2'] = MantenimientoForm(initial=model_to_dict(filtro))
                    template = get_template("adm_evento/modal/addevento.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            if action == 'edittipo':
                try:
                    data['id'] = request.GET['id']
                    data['filtro'] = filtro = TipoEvento.objects.get(pk=request.GET['id'])
                    data['form2'] = MantenimientoForm(initial=model_to_dict(filtro))
                    template = get_template("adm_evento/modal/addevento.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            if action == 'editperiodo':
                try:
                    data['title'] = u'Editar evento configurado'
                    data['periodo'] = periodo = PeriodoEvento.objects.get(pk=int(request.GET['id']))
                    # cantones = Canton.objects.filter(detalleperiodoevento__periodo=periodo,status=True)

                    cantones = Canton.objects.filter(pk__in=DetallePeriodoEvento.objects.values_list('canton_id').filter(periodo=periodo,status=True).distinct())
                    form = PeriodoEventoForm(initial={
                                                                     'periodo': periodo.periodo,
                                                                     'evento': periodo.evento,
                                                                     'tipo': periodo.tipo,
                                                                     'fechainicio': periodo.fechainicio,
                                                                     'fechainicioinscripcion': periodo.fechainicioinscripcion,
                                                                     'fechainicioconfirmar': periodo.fechainicioconfirmar,
                                                                     'fechafin': periodo.fechafin,
                                                                     'fechafininscripcion': periodo.fechafininscripcion,
                                                                     'fechafinconfirmar': periodo.fechafinconfirmar,
                                                                     'horainicio': periodo.horainicio,
                                                                     'horafin': periodo.horafin,
                                                                     'cuerpo': periodo.cuerpo,
                                                                     'iframemapa': periodo.iframemapa,
                                                                     'imagen': periodo.imagen,
                                                                     'portada': periodo.portada,
                                                                     'todos': periodo.todos,
                                                                     'permiteregistro': periodo.permiteregistro,
                                                                     'aplicaperiodo': periodo.aplicaperiodo,
                                                                     'validacupo': periodo.validacupo,
                                                                     'cupo': periodo.cupo,
                                                                     'descripcionbreve': periodo.descripcionbreve,
                                                                     'publicar': periodo.publicar,
                                                                     'tipoperfil': periodo.tipoperfil,
                                                                     'cantones': cantones,
                                                                     'link': periodo.link
                                                                     })

                    data['form'] = form
                    if not periodo.validacupo:
                        form.bloquear_cupo(0)
                    return render(request, "adm_evento/editperiodo.html", data)
                except Exception as ex:
                    pass

            if action == 'deltipo':
                try:
                    data['title'] = u'ELIMINAR TIPO DE EVENTO'
                    data['servicio'] = TipoEvento.objects.get(pk=request.GET['id'])
                    data['accion'] = 'deltipo'
                    data['url'] = '/adm_evento?action=configuraciones'
                    return render(request, 'adm_evento/delmantenimiento.html', data)
                except Exception as ex:
                    pass

            if action == 'delevento':
                try:
                    data['title'] = u'ELIMINAR EVENTO'
                    data['accion'] = 'delevento'
                    data['url'] = '/adm_evento?action=configuraciones'
                    data['servicio'] = Evento.objects.get(pk=request.GET['id'])
                    return render(request, 'adm_evento/delmantenimiento.html', data)
                except Exception as ex:
                    pass

            if action == 'delperiodo':
                try:
                    data['title'] = u'ELIMINAR PERIODO'
                    data['accion'] = 'delperiodo'
                    data['url'] = '/adm_evento'
                    data['servicio'] = PeriodoEvento.objects.get(pk=request.GET['id'])
                    return render(request, 'adm_evento/delmantenimiento.html', data)
                except Exception as ex:
                    pass

            if action == 'configuraciones':
                try:
                    data['title'] = u'Configuraciones'
                    data['eventos'] = Evento.objects.filter(status=True).order_by('id')
                    data['tipo'] = TipoEvento.objects.filter(status=True).order_by('id')
                    return render(request, "adm_evento/viewconfigurar.html", data)
                except Exception as ex:
                    pass

            if action == 'graficos':
                try:
                    data['title'] = 'Detalle de Inscritos'
                    options, departamento, desde, hasta, id, search, filtros, url_vars = request.GET.get('options', ''), request.GET.get('departamento', ''), request.GET.get('desde', ''),  request.GET.get('hasta', ''),  request.GET.get('id', ''),  request.GET.get('search', ''), Q(status=True), ''

                    data['periodoevento'] = periodoevento = PeriodoEvento.objects.get(pk=int(request.GET['id']))
                    querybase = RegistroEvento.objects.filter(periodo=periodoevento, status=True)
                    data['cantoneslist'] = cantones = Canton.objects.filter(status=True, id__in=querybase.values_list('canton_id',flat=True))

                    if desde:
                        data['desde'] = desde
                        filtros = filtros & Q(fecha_creacion__gte=desde)
                        url_vars += "&desde={}".format(desde)

                    if hasta:
                        data['hasta'] = hasta
                        filtros = filtros & Q(fecha_creacion__lte=hasta)
                        url_vars += "&hasta={}".format(hasta)

                    if departamento:
                        data['departamento'] = int(departamento)
                        filtros = filtros & (Q(departamento_id=departamento))
                        url_vars += '&departamento={}'.format(departamento)
                    if options:
                        data['options'] = int(options)
                        url_vars += "&options={}".format(options)

                    url_vars += '&action={}'.format(action)
                    data["url_vars"] = url_vars

                    querydetalle = querybase.filter(filtros)
                    listado = querydetalle.values_list('canton',flat=True).annotate(totcant=Count('canton'),  pendiente=Count(Case(When(estado_confirmacion=0, then=F('canton__id'),),distinct=True)), asistira=Count(Case(When(estado_confirmacion=1, then=F('canton__id'),),distinct=True)), noasistira=Count(Case(When(estado_confirmacion=2, then=F('canton__id'),),distinct=True)) ).values('canton__nombre', 'totcant', 'asistira', 'noasistira', 'pendiente').order_by('-totcant')
                    data['listado'] = listado
                    return render(request, 'adm_evento/graficas.html', data)
                except Exception as ex:
                    pass

            if action == 'vistaprevia':
                data['id'] = id = int((request.GET['id']))
                data['evento'] = evento = PeriodoEvento.objects.get(pk=id)

                data['title'] = u'{}'.format(evento.evento.nombre)
                return render(request, "adm_evento/vistaprevia.html", data)

            elif action == 'reporteinscritos':
                try:
                    evento = PeriodoEvento.objects.get(pk=request.GET['id'])
                    inscritos = evento.registroevento_set.filter(status=True)

                    __author__ = 'Unemi'

                    output = io.BytesIO()
                    workbook = xlsxwriter.Workbook(output)
                    ws = workbook.add_worksheet('reporte')
                    ws.set_column(0, 100, 60)

                    formatoceldagris = workbook.add_format(
                        {'align': 'center', 'border': 1, 'text_wrap': True, 'fg_color': '#B6BFC0'})
                    formatocelda = workbook.add_format(
                        {'align': 'left', 'border': 1, 'text_wrap': True})


                    ws.write(0, 0, 'EVENTO')
                    ws.merge_range('B1:F1', str(evento))
                    ws.write(1, 0, 'FACULTAD', formatoceldagris)
                    ws.write(1, 1, 'CARRERA', formatoceldagris)
                    ws.write(1, 2, 'NOMBRES', formatoceldagris)
                    ws.write(1, 3, 'APELLIDOS', formatoceldagris)
                    ws.write(1, 4, 'CEDULA', formatoceldagris)
                    ws.write(1, 5, 'DIRECCIÓN', formatoceldagris)
                    ws.write(1, 6, 'TELÉFONO', formatoceldagris)
                    ws.write(1, 7, 'CORREO INSTITUCIONAL', formatoceldagris)
                    ws.write(1, 8, 'CORREO PERSONAL', formatoceldagris)
                    ws.write(1, 9, 'TELÉFONO', formatoceldagris)
                    ws.write(1, 10, 'ESTADO', formatoceldagris)
                    cont=2
                    for inscrito in inscritos:
                        ws.write(cont,0, str(inscrito.inscripcion.coordinacion),formatocelda)
                        ws.write(cont,1, str(inscrito.inscripcion.carrera),formatocelda)
                        ws.write(cont,2, str(inscrito.participante.nombres),formatocelda)
                        ws.write(cont,3,'{} {}'.format(inscrito.participante.apellido1,inscrito.participante.apellido2 ),formatocelda)
                        ws.write(cont,4, str(inscrito.participante.identificacion()),formatocelda)
                        ws.write(cont,5, str(inscrito.participante.direccion_corta() ),formatocelda)
                        ws.write(cont,6, str(inscrito.participante.telefono ),formatocelda)
                        ws.write(cont,7, str(inscrito.participante.emailinst ),formatocelda)
                        ws.write(cont,8, str(inscrito.participante.email ),formatocelda)
                        ws.write(cont,9, str(inscrito.participante.telefono ),formatocelda)
                        ws.write(cont,10, str(inscrito.get_estado_confirmacion_display()),formatocelda)
                        cont+=1
                    workbook.close()
                    output.seek(0)
                    filename = 'reporte_inscritos.xlsx'
                    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    response['Content-Disposition'] = 'attachment; filename=%s' % filename
                    return response
                except Exception as ex:
                    pass

            elif action == 'reporteadmdoc':
                try:
                    evento = PeriodoEvento.objects.get(pk=request.GET['id'])
                    inscritos = evento.registroevento_set.filter(status=True)

                    __author__ = 'Unemi'

                    output = io.BytesIO()
                    workbook = xlsxwriter.Workbook(output)
                    ws = workbook.add_worksheet('reporte')
                    ws.set_column(0, 100, 60)

                    formatoceldagris = workbook.add_format(
                        {'align': 'center', 'border': 1, 'text_wrap': True, 'fg_color': '#B6BFC0'})
                    formatocelda = workbook.add_format(
                        {'align': 'left', 'border': 1, 'text_wrap': True})


                    ws.write(0, 0, 'EVENTO')
                    ws.merge_range('B1:F1', str(evento))
                    ws.write(1, 0, 'COD', formatoceldagris)
                    ws.write(1, 1, 'NOMBRES', formatoceldagris)
                    ws.write(1, 2, 'APELLIDOS', formatoceldagris)
                    ws.write(1, 3, 'CEDULA', formatoceldagris)
                    ws.write(1, 4, 'DIRECCIÓN', formatoceldagris)
                    ws.write(1, 5, 'TELÉFONO', formatoceldagris)
                    ws.write(1, 6, 'CORREO INSTITUCIONAL', formatoceldagris)
                    ws.write(1, 7, 'CORREO PERSONAL', formatoceldagris)
                    ws.write(1, 8, 'CARGO', formatoceldagris)
                    ws.write(1, 9, 'DEPARTAMENTO', formatoceldagris)
                    ws.write(1, 10, 'ESTADO', formatoceldagris)
                    cont=2
                    for inscrito in inscritos:
                        ws.write(cont,0, str(inscrito.id),formatocelda)
                        ws.write(cont,1, str(inscrito.participante.nombres),formatocelda)
                        ws.write(cont,2,'{} {}'.format(inscrito.participante.apellido1,inscrito.participante.apellido2 ),formatocelda)
                        ws.write(cont,3, str(inscrito.participante.identificacion()),formatocelda)
                        ws.write(cont,4, str(inscrito.participante.direccion_corta() ),formatocelda)
                        ws.write(cont,5, str(inscrito.participante.telefono ),formatocelda)
                        ws.write(cont,6, str(inscrito.participante.emailinst ),formatocelda)
                        ws.write(cont,7, str(inscrito.participante.email ),formatocelda)
                        if inscrito.participante.mi_cargo_actualadm():
                            plantilla = inscrito.participante.mi_cargo_actualadm()
                        else:
                            plantilla = inscrito.participante.mi_cargo_actual_docente()

                        ws.write(cont,8, str(plantilla.denominacionpuesto ) if plantilla else '',formatocelda)
                        ws.write(cont,9, str(plantilla.unidadorganica) if plantilla else '',formatocelda)
                        ws.write(cont,10, str(inscrito.get_estado_confirmacion_display()),formatocelda)
                        cont+=1
                    workbook.close()
                    output.seek(0)
                    filename = 'reporte_inscritos.xlsx'
                    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    response['Content-Disposition'] = 'attachment; filename=%s' % filename
                    return response
                except Exception as ex:
                    pass


            elif action == 'cargarmuestra':
                try:
                    data['id'] = request.GET['id']
                    data['filtro'] = filtro = PeriodoEvento.objects.get(pk=request.GET['id'])
                    data['form2'] = CargarMuestraForm()
                    template = get_template("noticias/cargarmuestra.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            return HttpResponseRedirect(request.path)
        else:
            data['title'] = u'Eventos de la institución'

            url_vars = ''
            filtro = Q(status=True)
            search = None
            ids = None
            tipo = None

            if 't' in request.GET:
                if request.GET['t'] != '0':
                    tipo = request.GET['t']
            if 's' in request.GET:
                if request.GET['s'] != '':
                    search = request.GET['s']

            if search:
                filtro = filtro & (Q(evento__descripcion__icontains=search))
                url_vars += '&s=' + search

            if tipo:
                filtro = filtro & Q(tipoperfil=int(tipo))
                url_vars += '&t=' + tipo

            periodo = PeriodoEvento.objects.filter(filtro).order_by('-id')

            paging = MiPaginador(periodo, 20)
            p = 1
            try:
                paginasesion = 1
                if 'paginador' in request.session:
                    paginasesion = int(request.session['paginador'])
                if 'page' in request.GET:
                    p = int(request.GET['page'])
                    p = int(request.GET['page'])
                else:
                    p = paginasesion
                try:
                    page = paging.page(p)
                except:
                    p = 1
                page = paging.page(p)
            except:
                page = paging.page(p)
            request.session['paginador'] = p
            data['paging'] = paging
            data['rangospaging'] = paging.rangos_paginado(p)
            data['page'] = page
            data['search'] = search if search else ""
            data['t'] = int(tipo) if tipo else ''
            data["url_vars"] = url_vars
            data['ids'] = ids if ids else ""
            data['periodos'] = page.object_list
            data['tipo'] = TipoEvento.objects.filter(status=True)
            data['email_domain'] = EMAIL_DOMAIN
            return render(request, 'adm_evento/view.html', data)