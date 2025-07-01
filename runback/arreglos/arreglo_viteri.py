# coding=utf-8
#!/usr/bin/env python

import os
import sys
from datetime import datetime, time, timedelta
from time import strftime

# SITE_ROOT = os.path.dirname(os.path.realpath(__file__))
import xlrd
import xlsxwriter
from openpyxl import load_workbook
# from pywhatkit import sendwhatmsg_instantly


YOUR_PATH = os.path.dirname(os.path.realpath(__file__))
SITE_ROOT = os.path.dirname(os.path.dirname(YOUR_PATH))
SITE_ROOT = os.path.join(SITE_ROOT, '')
your_djangoproject_home = os.path.split(SITE_ROOT)[0]
sys.path.append(your_djangoproject_home)

from django.core.wsgi import get_wsgi_application
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "settings")
application = get_wsgi_application()
from datetime import datetime, timedelta
from django.db import transaction
from sga.models import Materia, Clase, Leccion, LeccionGrupo, AsistenciaLeccion
from sga.funciones import variable_valor, validar_ldap_reseteo, convertir_hora
from sga.models import *
from sagest.models import *
from postulate.models import *
from moodle.models import UserAuth
from inno.models import *
from helpdesk.models import BodegaKardex


from settings import MEDIA_ROOT
#directory = os.path.join(MEDIA_ROOT, 'reportes', 'reporte_requisitos.xls')

# actividadsecuencial = ActividadSecuencialTH.objects.filter(status=True)
# print('Inicia')
# cont=0
# for actividad in actividadsecuencial:
#     gestionproducto = GestionProductoServicioTH.objects.filter(producto=actividad.servicios,gestion=actividad.gestion,status=True)
#     if not gestionproducto:
#         gestionproducto = GestionProductoServicioTH(producto=actividad.servicios,
#                                                                    gestion=actividad.gestion, status=True)
#         gestionproducto.save()
#     else:
#         gestionproducto = gestionproducto[0]
#     actividad.producto = gestionproducto
#     actividad.save()
#     cont += 1
#     print('registo %s de %s' % (cont,actividadsecuencial.count()))
# print('fin, total migrados %s de %s' % (cont,actividadsecuencial.count()))

#     if producto.productointermedio:
#         if not ProductoIntermedioTH.objects.filter(descripcion=(producto.productointermedio).strip().upper()).exists():
#             inter = ProductoIntermedioTH(descripcion=(producto.productointermedio).strip().upper())
#             inter.save()
#             cont += 1
# print('fin, total migrados %s de %s' % (cont,productos.count()))
#

def actualizar_secciones_productos():
    secciones = SeccionDepartamento.objects.filter(status=True)
    contp=0
    print("Inicia migracion")
    for seccion in secciones:
        productos = seccion.productoservicioseccion_set.filter(status=True,activo=True)
        print("inicia gestion: %s" % seccion)

        for pro in productos:
            gest = GestionPlanificacionTH.objects.filter(gestion=seccion,status=True)
            for ge in gest:
                if not GestionProductoServicioTH.objects.filter(gestion=ge,producto=pro.producto,status=True):
                    migrar = GestionProductoServicioTH(gestion=ge,producto=pro.producto)
                    migrar.save()
                    contp+=1
    print("fin, total productos migrados: %s" %contp)
    conta=0
    actividades = ActividadSecuencialTH.objects.filter(status=True)
    print("inicia migrar actividades")
    for acti in actividades:
        acti.servicios=acti.servicio.producto
        acti.save()
        conta+=1
    print("fin, total actividades migradas: %s" %conta)


# archivo_ = 'areas_conocimiento'
# url_archivo = "{}/media/{}.xlsx".format(SITE_STORAGE, archivo_)
# workbook = load_workbook(filename=url_archivo, read_only=False)
# sheet = workbook[workbook.sheetnames[0]]
# linea = 1
# col_codcaamplio = 1
# col_codciamplio = 2
# col_amplio = 3
# col_codcaespecifico = 4
# col_codciespecifico = 5
# col_especifico = 6
# col_codcadetallado = 7
# col_codcidetallado = 8
# col_detallado = 9
# for rowx in range(2, sheet.max_row + 1):
#     cocaamplio = (sheet.cell(row=rowx,column=col_codcaamplio).value)
#     codciamplio = sheet.cell(row=rowx,column=col_codciamplio).value
#     amplio = (sheet.cell(row=rowx,column=col_amplio).value).upper().strip()
#     codcaespecifico = sheet.cell(row=rowx,column=col_codcaespecifico).value
#     codciespecifico = sheet.cell(row=rowx,column=col_codciespecifico).value
#     especifico = (sheet.cell(row=rowx,column=col_especifico).value).upper().strip()
#     codcadetallado = sheet.cell(row=rowx,column=col_codcadetallado).value
#     codcidetallado = sheet.cell(row=rowx,column=col_codcidetallado).value
#     detallado = (sheet.cell(row=rowx,column=col_detallado).value).upper().strip()
#     areaconocimiento = AreaConocimientoTitulacion.objects.filter(Q(nombre=amplio)|Q(codigo=cocaamplio))
#     subareaconocimiento = SubAreaConocimientoTitulacion.objects.filter(nombre=especifico)
#     subareadetallado = SubAreaEspecificaConocimientoTitulacion.objects.filter(nombre=detallado)
#
#     if areaconocimiento:
#         areaconocimiento=areaconocimiento[0]
#         #areaconocimiento.codigo=str(codciamplio)
#         areaconocimiento.codigocaces=str(cocaamplio)
#         areaconocimiento.migrado=True
#         areaconocimiento.save()
#         print (areaconocimiento)
#     else:
#         print('No existe area conocimiento: %s' % amplio)
#
#     if subareaconocimiento:
#         subareaconocimiento = subareaconocimiento[0]
#         subareaconocimiento.codigo = str(codciespecifico)
#         subareaconocimiento.codigocaces = str(codcaespecifico)
#         subareaconocimiento.migrado = True
#         subareaconocimiento.save()
#         print(subareaconocimiento)
#     else:
#         print('No existe subarea conocimiento: %s' % especifico)
#
#     if subareadetallado:
#         subareadetallado = subareadetallado[0]
#         #subareadetallado.codigo = str(codcidetallado)
#         subareadetallado.codigocaces = str(codcadetallado)
#         subareadetallado.migrado = True
#         subareadetallado.save()
#         print(subareadetallado)
#     else:
#         print('No existe subarea detallada conocimiento: %s' % detallado)

    #print(cocaamplio,codciamplio,amplio,codcaespecifico,codciespecifico,especifico)


#PROCESO COMPLETAR BECAS DE ESTUDIANTES
# becatotal = BecaSolicitud.objects.values('inscripcion_id').filter(periodo_id=317,status=True)
# cantidad_limite_becados = BecaPeriodo.objects.get(periodo_id=317,status=True)
# cuentas = CuentaBancariaPersona.objects.values('persona_id').filter(estadorevision__in=[1,2,5])
# becaincluir = BecaSolicitud.objects.values('inscripcion_id').filter(status=True,becaaceptada=2, inscripcion__persona_id__in=(cuentas)).exclude(periodo_id=317)
# cantidad_estudiantes_becados = BecaSolicitud.objects.values('id').filter(status=True,periodo_id=317).exclude(
#                     becaaceptada=3).count()
# preinscriones = PreInscripcionBeca.objects.filter( inscripcion_id__in=becaincluir,periodo_id=317).order_by('orden')
# # while cantidad_estudiantes_becados < cantidad_limite_becados.limitebecados:
# print("inicia proceso solicitar beca")
# contingres=0
# for prei in preinscriones:
#     if cantidad_estudiantes_becados < cantidad_limite_becados.limitebecados:
#         if not BecaSolicitud.objects.filter(status=True,periodo_id=317, periodocalifica_id=224,inscripcion=prei.inscripcion).exists():
#             becado = BecaSolicitud(inscripcion=prei.inscripcion,
#                                    becatipo=prei.becatipo,
#                                    periodo_id=317,
#                                    periodocalifica_id=224,
#                                    estado=1,
#                                    tiposolicitud=2,
#                                    observacion=f'SEMESTRE REGULAR {prei.periodo.nombre}')
#             becado.save(usuario_id=1)
#             recorrido = BecaSolicitudRecorrido.objects.filter(solicitud=becado, estado=1).first()
#             prei.seleccionado = True
#             prei.save(usuario_id=1)
#             if recorrido is None:
#                 recorrido = BecaSolicitudRecorrido(solicitud=becado,
#                                                    observacion="SOLICITUD AUTOMÁTICA",
#                                                    estado=1)
#                 recorrido.save(usuario_id=1)
#                 recorrido = BecaSolicitudRecorrido.objects.filter(solicitud=becado, estado=4).first()
#                 # REGISTRO EN ESTADO DE REVISION
#                 if recorrido is None:
#                     becado.estado = 4
#                     becado.save(usuario_id=1)
#                     recorrido = BecaSolicitudRecorrido(solicitud=becado,
#                                                        observacion="EN REVISION",
#                                                        estado=4)
#                     recorrido.save(usuario_id=1)
#                     # log(u'Agrego recorrido de beca: %s' % recorrido, request, "add")
#
#                 recorrido = BecaSolicitudRecorrido.objects.filter(solicitud=becado, estado=2).first()
#                 if recorrido is None:
#                     becado.estado = 2
#                     becado.becaaceptada = 1
#                     becado.save(usuario_id=1)
#                     recorrido = BecaSolicitudRecorrido(solicitud=becado,
#                                                        observacion="PENDIENTE DE ACEPTACIÓN O RECHAZO",
#                                                        estado=2)
#                     recorrido.save(usuario_id=1)
#                     # log(u'Agrego recorrido de beca: %s' % recorrido, request, "add")
#
#
#                     cantidad_estudiantes_becados = BecaSolicitud.objects.values('id').filter(status=True,
#                                                                                              periodo_id=317).exclude(
#                         becaaceptada=3).count()
#                     contingres+=1
#                     print(contingres,'registrado',becado)
#     else:
#         break
#
# becas = BecaSolicitud.objects.filter(periodo_id=224,status=True,becaaceptada=1)
# print("total seleccionados:",becas.count())
# aData = {}
# cont = 0
# cont2 = 0
# print("Inicia proceso")
# for becasolicitud in becas:
#     if BecaAsignacion.objects.filter(solicitud__inscripcion=becasolicitud.inscripcion, status=True).exclude(solicitud__periodo_id=224).exists():
#         becatipoconfiguracion = becasolicitud.obtener_configuracionbecatipoperiodo()
#         montobeca = becatipoconfiguracion.becamonto
#         meses = becatipoconfiguracion.becameses
#         montomensual = becatipoconfiguracion.monto_x_mes()
#         becaasignacion = BecaAsignacion.objects.filter(solicitud=becasolicitud, status=True).first()
#         if not becaasignacion:
#             beca = BecaAsignacion(solicitud=becasolicitud,
#                                   montomensual=montomensual,
#                                   cantidadmeses=meses,
#                                   montobeneficio=montobeca,
#                                   fecha=datetime.now().date(),
#                                   activo=True,
#                                   grupopago=None,
#                                   tipo=2,
#                                   notificar=True,
#                                   estadobeca=None,
#                                   infoactualizada=False,
#                                   cargadocumento=True)
#             beca.save()
#             becasolicitud.tiposolicitud = 2
#             becasolicitud.becaaceptada = 2
#             becasolicitud.becaasignada = 2
#             becasolicitud.save()
#             cont+=1
#             print(cont," Aceptado a: ",str(becasolicitud.inscripcion.persona))
# print("Total creados: ",cont, "Total actualizados: ",cont2)

# gestiones = GestionProductoServicioTH.objects.filter(status=True)
# for gestion in gestiones:
#     producto = ProductoServicioSeccion.objects.filter(status=True, producto=gestion.producto, seccion=gestion.gestion.gestion)
#     if producto:
#         print('Encontrado',producto.count())
#         producto=producto[0]
#         gestion.productoseccion=producto
#         gestion.save()
#     else:
#         print('No se encontró producto')

# actualizar tribuna con  modelo evaluativo
def extraer_cargo():
   distributivo = DistributivoPersonaHistorial.objects.filter(persona_id=30069).order_by('-fecha_creacion')
   for dis in distributivo:
    print(dis.unidadorganica, dis.denominacionpuesto,dis.fecha_creacion )

def actualizar_productos():
    gestiones = GestionPlanificacionTH.objects.filter(cabecera__periodo_id=9,status=True)
    actualizados=0
    nuevos=0
    for gestion in gestiones:
        productos=gestion.gestion.productoservicioseccion_set.filter(status=True,activo=True)
        for producto in productos:
            produsecc = GestionProductoServicioTH.objects.filter(productoseccion=producto,gestion=gestion,status=True)
            if produsecc:
                actualizados+=1
                produsecc=produsecc[0]
                produsecc.activoseccion=True
                produsecc.producto=producto.producto
            else:
                nuevos+=1
                produsecc=GestionProductoServicioTH(productoseccion=producto,
                                                    gestion=gestion,
                                                    activoseccion=True,
                                                    producto=producto.producto)
            produsecc.save()
    print('actualizados: ',actualizados, 'Creados: ',nuevos)


def definir_mejores_puntuados():
    cont=0
    partidas = Partida.objects.filter(convocatoria__status=True,status=True)
    for partida in partidas:
        for postulante in partida.personaaplicarpartida_set.filter(status=True, estado__in=[1, 4, 5]).order_by(
                        '-nota_final_meritos')[:partida.convocatoria.nummejorespuntuados]:
            postulante.esmejorpuntuado=True
            cont+=1
            postulante.save()
            print(cont, 'mejor puntuado', postulante,)

def retirar_duplicados_practicas():
    preinscripciones = DetallePreInscripcionPracticasPP.objects.filter(usuario_creacion_id=1,estado=2,
                                                                     itinerariomalla_id__in=[126,128,130,125,129,127,137,173,175,176,174,172,136,195,135,134,193,196,230,194],status=True)
    print(preinscripciones.count())
    total=0
    for preinscripcion in preinscripciones:
        verificar = DetallePreInscripcionPracticasPP.objects.filter(inscripcion=preinscripcion.inscripcion,
                                                                    status=True,
                                                                    itinerariomalla=preinscripcion.itinerariomalla, estado=1).exclude(id=preinscripcion.id)
        verificar.update(status=False)
        if verificar.count()>0:
            total=total+verificar.count()
            print(verificar)
    print(total)

def comprobar_baja():
    activos=[21634558, 21625245, 18654532, 7237823, 7237748, 7237845, 7236501, 7235851, 7237763, 762925,
             7236918, 7236654, 7236373, 17534151, 7237792, 23252114, 7237800, 7235633, 23252131, 18886108,
             7236484, 7237242, 21634780, 21625225, 7237860, 19427673, 7236561, 7236477, 7236392, 7235754,
             17532662, 30983885, 2671417, 17533454, 7237864, 7237588, 868437, 31584810, 7237790, 7235484,
             7236582, 7236531, 7236236, 7235531, 28355857, 23252241, 21634798, 7237818, 7237682, 7237248,
             7236807, 7236557, 7235764, 21634564, 14367513, 14367490, 7237659, 7237188, 7236360, 7236688,
             7236273, 868024, 7560362, 7237625, 27167528, 2493968, 23232912, 19427618, 7237871, 7237843, 7236997,
             7236491, 19427645, 23232906, 7237749, 7237208, 7235865, 21634772, 21634522, 21625224, 7236980, 7237896,
             28007862, 7237692, 17532650, 7237476, 7237401, 7237360, 7237351, 7237321, 7237111, 7237104, 7236659,
             21625048, 19427715, 19427583, 18654636, 18654635, 17228297, 7236671, 7236638, 7236190, 17534147, 17533445,
             7237060, 7560287, 7237107, 7237054, 7237043, 7236514, 29947992, 28355840, 21634814, 21634542, 21625262,
             14367481, 28330693, 916069, 916068, 916066, 916065, 916064, 916062, 916060, 916059, 7237653, 7237642,
             7237628, 7237605, 7237604, 7237586, 7237482, 7237283, 7237281, 7237274, 7237267, 7237259, 7237254, 7237253,
             7237244, 7237240, 7237235, 7237234, 7237232, 7237231, 7237230, 7237222, 7237216, 7237215, 7237213, 7237200,
             7237001, 7236436, 7236428, 7236394, 29948278, 29948086, 29948066, 29948028, 21634813, 21634812, 21634809,
             21634808, 21634807, 21634806, 21634796, 21634795, 21634794, 21634565, 21634562, 21634559, 21634556,
             21634551, 21634550, 21634549, 21634547, 21634546, 21625261, 21625258, 21625257, 21625246, 21625242,
             21625241, 19427601, 14367600, 7237759, 31586751, 28007860, 7559894, 7237103, 7237018, 7236506, 32060170,
             23252277, 23252172, 7237634, 23252235, 762909, 7237179, 17532632, 7237867, 19427597, 7235587, 7237567,
             7237553, 7236109, 7560364, 29948274, 21634529, 21634524, 7237765, 7235606, 14367635, 14367498, 23252259,
             29948052, 29948044, 29948037, 29948032, 29948017, 23252253, 7237878, 7237852, 7237804, 7237799, 7237648,
             7237313, 7237311, 7237165, 7236849, 7236837, 7236578, 7236504, 7236389, 7236290, 7236233, 7236228, 7236223,
             7236211, 7236207, 7236206, 7236201, 7236200, 7236046, 7236044, 7236039, 7236021, 7236020, 7236009, 7235998,
             7235993, 7235989, 7235986, 7235974, 7235968, 7235963, 7235955, 7235927, 7235909, 7235902, 7235774, 7235773,
             7235749, 7235744, 7235706, 7235577, 7237839, 7237272, 7559882, 7237539, 7237325, 17533486, 7237282,
             7237072, 28094338, 28007866, 7237750, 7237838, 23252224, 21625243, 23252244, 7237889, 7236209, 7237095,
             7237207, 7237873, 7237372, 7237017, 7235848, 17532375, 7236622, 7236193, 7236635, 21625106, 7237788,
             7237774, 7237199, 7237189, 7237053, 7237051, 7237032, 7236612, 7236611, 7236551, 7236383, 7235868, 3151039,
             2493975, 17532920, 7236434, 7236330, 19427620, 17533481, 17533469, 17533465, 7237396, 7237381, 7237343,
             7237211, 7236443, 21634759, 21634752, 21634750, 21634749, 21634748, 21634746, 21634745, 21634744, 21634740,
             21634739, 21634738, 21634737, 21634736, 21634733, 21634732, 21634731, 21634729, 21634728, 21634727,
             21634726, 21634725, 21634724, 21634723, 21634722, 21634721, 21634720, 21634719, 21634718, 21634715,
             21634714, 21634713, 21634712, 21634711, 21634710, 21634709, 21634708, 21634707, 21634706, 21634705,
             21634704, 21634703, 21634702, 21634701, 21634700, 21634699, 21634694, 21634691, 21634680, 18654450,
             7235584, 23232910, 7237034, 7235505, 7237143, 7237074, 14367616, 14367499, 7560280, 7237068, 7236343,
             7236000, 7235928, 7235917, 7235685, 7235615, 19427721, 7237876, 7560398, 7236954, 7236909, 7235498,
             17533458, 7236928, 7237404, 7237197, 7237099, 7236962, 7236439, 7235864, 7236620, 7235861, 17534023,
             17533436, 29948212, 7237265, 7237176, 17532649, 7237190, 7235730, 7236902, 7236839, 7236790, 7236788,
             7236787, 7236784, 21634523, 21634521, 21634520, 21625505, 21625500, 21625498, 21625488, 21625486, 21625483,
             21625482, 21625481, 21625478, 21625476, 21625473, 21625468, 21625467, 21625466, 21625248, 21625238,
             21625233, 21625231, 21625227, 21625226, 21625223, 21625222, 21625220, 21625216, 21625213, 14367531,
             7236521, 7237738, 7236421, 7235684, 7236533, 7236250, 17532633, 7560051, 7559726, 7237011, 7236028,
             7236594, 7237629, 7236426, 17533439, 7237299, 19427722, 7235786, 7235785, 7235775, 7235770, 7235767,
             7235761, 7235757, 7235747, 7235742, 7235693, 7235518, 19427661, 17534153, 17533662, 7237709, 7236560,
             21634805, 21634803, 21634793, 21634790, 21634789, 21634788, 21634787, 21634783, 21634781, 21634778,
             21634777, 21634775, 21634773, 21634771, 21634770, 21634768, 21634555, 21634553, 21634543, 21634541,
             21634540, 21634539, 21634538, 21634537, 21634533, 21634532, 21634531, 21634528, 21634527, 21634525,
             7235850, 7237798, 7236249, 7237061, 23252168, 17532376, 28330686, 18650393, 9107521, 9107516, 7560448,
             7560447, 7560443, 7560442, 7560439, 7560438, 7560433, 7237892, 7237436, 7237434, 7237426, 7237417, 7237414,
             7237270, 7237236, 7237031, 7237025, 7237013, 7236943, 7236440, 7236358, 7236262, 7236177, 7235836, 7235611,
             3151035, 30752336, 30752281, 30355577, 28355894, 28355842, 28355841, 28102511, 23252227, 23252181, 7236780,
             7236779, 7236775, 7236774, 7236773, 7236771, 7236766, 7236284, 7236279, 7236191, 7236016, 7235948, 7235876,
             7235517, 7235506, 30355576, 18304087, 17534148, 17532657, 11741, 17532635, 7559887, 7236739, 7236675,
             7236632, 7235645, 7235634, 7237786, 23232904, 762911, 762910, 7237812, 7237773, 7237599, 7237122, 7236280,
             7237856, 7237802, 29948027, 29948007, 7236349, 23252177, 7237652, 7237155, 7237138, 7236455, 21625463,
             7237276, 7237159, 7237010, 7237007, 29948172, 29948047, 28355813, 27977081, 18654425, 7237045, 7236628,
             7237832, 7235789, 7237791, 23252146, 7237187, 7237336, 7237271, 7237142, 7237065, 7236835, 7236651,
             7235631, 21634779, 21634530, 21625469, 21625228, 19427733, 7237268, 7237158, 7236569, 7237615, 23252133,
             7560567, 7237617, 7237610, 7236858, 7236595, 7236473, 7236248, 7236235, 7235822, 7235819, 7235809, 7235808,
             14367508, 7236335, 5490445, 915868, 915853, 915845, 915830, 915821, 915812, 915803, 915794, 915785, 915778,
             915770, 915761, 915753, 915745, 915736, 915720, 915705, 915697, 7237883, 7237740, 7237739, 7237719,
             7237717, 7237698, 7237691, 7237683, 7237672, 7237666, 7237663, 7237646, 7237643, 7237635, 7237624, 7237618,
             7237616, 7237595, 7237551, 7237118, 7237041, 7236978, 7236850, 7236678, 7236646, 7236356, 29948149,
             29948020, 2692594, 2692592, 2692591, 14289053, 7559538, 7560054, 7237560, 7237658, 7235853, 7235585,
             28007859, 7237370, 7237364, 7237353, 7237185, 7236281, 7235729, 7236970, 7236364, 7236225, 23252260,
             7237847, 17533617, 17533615, 916074, 916070, 916063, 916061, 7237737, 7237734, 7237730, 7237707, 7237706,
             7237700, 7237688, 7237680, 7237678, 7237661, 7237640, 7237632, 7237626, 7237613, 7237611, 7237584, 7237579,
             7237350, 7236906, 7236869, 7236438, 7236285, 7236231, 7235594, 21634635, 21634625, 21634622, 21634599,
             21634596, 21634580, 21634575, 21634390, 21634382, 21634380, 21634374, 21634361, 21634357, 21634333,
             21634332, 21634325, 21634320, 21634318, 21625325, 14367522, 17533659, 17533656, 17533652, 17533657,
             7237581, 23252233, 7559814, 7237264, 7236589, 7235871, 29948214, 23252121, 762914, 7237746, 7237275,
             7237219, 7237133, 916054, 916047, 916040, 916033, 916026, 916018, 916010, 916003, 915994, 915986, 915978,
             915971, 915964, 915957, 915950, 915942, 915934, 915928, 915920, 915914, 915908, 915901, 915893, 915884,
             915876, 915861, 915712, 915685, 864508, 7237813, 7237742, 7237735, 7237733, 7237732, 7237726, 7237721,
             7237714, 7237712, 7237710, 7237705, 7237695, 7237687, 7237685, 7237681, 7237677, 7237673, 7237115, 7236991,
             7236984, 7236938, 7236932, 7236479, 7236254, 7236224, 7236219, 7236185, 3151047, 2493973, 7237862,
             28007851, 7237884, 7237175, 7236445, 7236351, 14367493, 7237841, 7237620, 31586745, 7237701, 7237679,
             7237577, 7237255, 7237241, 2493964, 23252202, 19700308, 7237636, 7237546, 7559328, 7237781, 7235527,
             7236950, 7236187, 28007852, 7237713, 7237881, 7237868, 7237037, 7559593, 7559592, 7235715, 7236941,
             7236549, 7236400, 19427617, 19427616, 19427613, 19427582, 7236644, 7236192, 7560045, 7560033, 7559868,
             5495296, 5495287, 3157121, 3157120, 3157119, 3157115, 3157113, 3157109, 3157107, 7560368, 7560353, 7560351,
             7237335, 7237166, 7237047, 7236903, 7236895, 7236888, 7236881, 7236872, 7236864, 7236610, 7235501, 7235500,
             23252273, 23252251, 21634784, 21634534, 21625479, 21625229, 14367660, 17533543, 17533533, 7237815, 7237593,
             3151042, 7237400, 7237390, 7237362, 7237358, 7237280, 7237279, 7237266, 7237262, 7237224, 7237218, 7237205,
             7237203, 7236469, 7236346, 7236334, 7236329, 7236313, 7236308, 7236304, 7236302, 7236296, 7236295, 7236041,
             7235960, 7235922, 21634747, 21634743, 21634735, 21634734, 21634672, 21634509, 21634507, 21634492, 18654630,
             18654619, 18654614, 18654607, 18654556, 18654554, 18654550, 18654548, 18654542, 18654538, 18654536,
             18654534, 18654530, 18654528, 18654520, 18654511, 18654508, 18654491, 18654487, 18654486, 18654482,
             18654480, 18654479, 18654478, 18654462, 18654460, 18654459, 18654447, 18654446, 18654444, 18654441,
             18654439, 18654438, 18654433, 18654428, 18654426, 7237649, 3157114, 3157105, 7237402, 7237383, 7237382,
             7237377, 7237369, 7237352, 7237347, 7237338, 7237323, 7237322, 7237308, 7237304, 7237300, 7237206, 7236687,
             7237807, 7237305, 7237561, 7237842, 7236345, 14367538, 7560350, 7559825, 7237718, 7237163, 7235994,
             23252116, 21625472, 14367487, 14367475, 7237745, 7236894, 7236423, 7236203, 30355601, 30355580, 19427730,
             7237794, 26585854, 21625249, 7237117, 7236684, 7236647, 7236189, 7560358, 7560076, 7559741, 7559662,
             7236802, 7236795, 7236792, 7235728, 21634816, 21634566, 21625259, 7560273, 7560071, 7560060, 7559881,
             7559879, 7559842, 7559838, 7559837, 7237485, 7237389, 7237339, 7237303, 7237291, 7236923, 7236870, 7236862,
             7236855, 7236854, 7236853, 7236846, 7236844, 7236843, 7236836, 7236834, 7236747, 7236708, 7236683, 7236669,
             7236656, 7236636, 7236524, 7236232, 7236226, 7236217, 7236216, 7236208, 7236106, 7236070, 7235881, 7235874,
             7235856, 7235719, 7235680, 7235671, 7235658, 7235644, 7235642, 7235624, 7235579, 7235576, 17532398,
             7237194, 7237225, 7236220, 7235622, 7235600, 7237769, 7236934, 7237227, 7237882, 28007861, 7237228,
             7235699, 29948187, 28355855, 28007831, 7237545, 7237137, 7236915, 7236889, 7236670, 7235824, 7235727,
             14367564, 7237162, 7237056, 7236899, 7236558, 7236556, 7236553, 7236550, 7236547, 7236546, 7236545,
             7236539, 7236537, 7236536, 7236526, 7236519, 7236513, 7236511, 7236498, 7236490, 7236487, 7236483, 7236482,
             7236480, 7236475, 7236467, 7236458, 7236454, 7236450, 7236447, 7236446, 7236442, 7236435, 7236432, 7236427,
             7236425, 7236424, 7236420, 7236418, 7236410, 7236402, 7236399, 7236398, 7236395, 7236393, 7236390, 7236387,
             7236386, 7236385, 7236381, 7236377, 7236374, 7236372, 7236365, 7236362, 7236357, 7236354, 7236333, 7236011,
             21634766, 21625209, 21625206, 17533950, 3151041, 7237036, 7235573, 7237583, 7237183, 7236538, 7237716,
             7237552, 19427653, 7560290, 916058, 916057, 7237668, 7237665, 7237657, 7237651, 7237647, 7237638, 7237637,
             7237633, 7237630, 7237619, 7237612, 7237609, 7237602, 7237597, 7237590, 7237589, 7237582, 7237575, 7237329,
             7236306, 7236301, 7236286, 7235958, 7235549, 28355831, 21634663, 21634417, 21634416, 21634415, 21634413,
             21634412, 21634411, 21634410, 21634409, 21634408, 21634407, 21634406, 21634405, 21634404, 21634403,
             21634402, 21634401, 21634400, 21634399, 21634398, 21634397, 21634396, 21634395, 21634394, 21634393,
             21634392, 14367657, 14367656, 14367655, 14367648, 14367647, 14367643, 14367642, 14367529, 14367518,
             14367489, 14367470, 14367464, 21625215, 7237600, 7236548, 7236406, 7560607, 7560606, 7560604, 7560602,
             7560601, 7560598, 7560594, 7560592, 7560591, 7560587, 7560582, 7560581, 7560578, 7560574, 7560573, 7560569,
             7560568, 7560565, 5495278, 5495249, 3157123, 3157122, 3157118, 3157116, 3157112, 3157110, 3157108, 3157106,
             3157104, 7237805, 7237669, 7237392, 7237344, 7237326, 7237314, 7237309, 7237307, 7237296, 7237295, 7236900,
             7236883, 7236874, 7236816, 7236810, 7236764, 7236761, 7236759, 7236756, 7236750, 7236748, 7236743, 7236740,
             7236735, 7236733, 7236730, 7236721, 7236715, 7236712, 7236705, 7236456, 7236367, 28355849, 21634785,
             21634776, 21634535, 21634526, 21625471, 21625230, 21625221, 14367548, 14367546, 14367539, 14367526,
             14367523, 14367507, 19427623, 19427600, 19427594, 7236583, 23252226, 7560274, 7560545, 7237566, 7236959,
             7236930, 7236234, 7235570, 7237607, 7236494, 7237888, 19427614]
    cont = 0
    print('NRO - FECHA - OFICIO - TIPO_BAJA - ESTADO - COD_INTERNO - COD_GOBIERNO - RESPONSABLE - CATEGORIA')
    for i in activos:
        activo = ActivoFijo.objects.filter(Q(codigogobierno=str(i)) | Q(codigointerno=str(i)),status=True).first()
        if activo:
            # print(cont,activo)
            if activo.statusactivo==2:
                cont += 1
                detalle = DetalleBajaActivo.objects.filter(seleccionado=True,status=True,codigobaja__status=True,activo=activo).first()
                print('%s - %s - %s - %s - %s - %s - %s - %s - %s' % (detalle.codigobaja.numero,detalle.codigobaja.fecha,
                                                                      detalle.codigobaja.oficio,
                      detalle.codigobaja.tipobaja,detalle.codigobaja.estado,
                      activo.codigointerno,activo.codigogobierno,activo.responsable, activo.catalogo.grupo))

def actualizar_constatacion():
    activos = DetalleConstatacionFisica.objects.filter(status=True,codigoconstatacion__periodo__isnull=False)
    secuencial = ConstatacionFisica.objects.filter(periodo__isnull=False)
    secuencial.update(numero=0,estado=1)
    for activo in activos:
        activo.ubicacionanterior=activo.activo.ubicacion
        activo.save()
        print("Ub. anterior:",activo.ubicacionanterior,"Ub. actual:", activo.ubicacionbienes)

def verificar_pregunta(pazsalvo):
    from sagest.models import ActivoFijo
    pazsalvos = PazSalvo.objects.filter(status=True,estado=1)
    for pazsalvo in pazsalvos:
        funcionario = pazsalvo.persona
        activos = ActivoFijo.objects.filter(Q(responsable=funcionario) | Q(custodio=funcionario),status=True).exists()
        cargos = funcionario.mis_cargos().exclude(denominacionpuesto=pazsalvo.cargo).exists()
        if not (not activos or cargos):
            pregunta = DetallePazSalvo.objects.filter(persona_id=1204, pazsalvo=pazsalvo, status=True)
            if pregunta:
                pregunta.update(respondio = False)

def marcar_masivo():
    archivo_ = 'subir_marcadas'
    url_archivo = "{}/media/{}.xlsx".format(SITE_STORAGE, archivo_)
    workbook = load_workbook(filename=url_archivo, read_only=False)
    sheet = workbook[workbook.sheetnames[1]]
    linea = 1
    col_cedula = 2
    for rowx in range(2, sheet.max_row + 1):
        cedula = (sheet.cell(row=rowx,column=col_cedula).value)
        if cedula:
            servidor = Persona.objects.filter(Q(cedula=cedula)|Q(pasaporte=cedula),status=True)
            if servidor:
                servidor = servidor[0]
                hoy = datetime.now()
                fecha = date(2024, 2, 7)
                secuencia = 0
                jornada = HistorialJornadaTrabajador.objects.filter(status=True, persona=servidor,
                                                                    fechainicio__lte=hoy,
                                                                    fechafin__isnull=True).first()
                if jornada:
                    detalles = jornada.jornada.detalle_jornada()
                    detalles = detalles.filter(dia=2)
                    logdia = LogDia.objects.filter(status=True, fecha=fecha, persona=servidor).first()
                    if not logdia:
                        logdia = LogDia(fecha=fecha, persona=servidor, jornada=jornada.jornada)
                        logdia.save()
                    marcadas_old = LogMarcada.objects.filter(logdia=logdia)
                    marcadas_old.delete()
                    for deta in detalles:
                        fin = datetime(fecha.year, fecha.month, fecha.day, deta.horafin.hour, deta.horafin.minute,
                                       deta.horafin.second)
                        inicio = datetime(fecha.year, fecha.month, fecha.day, deta.horainicio.hour,
                                          deta.horainicio.minute, deta.horainicio.second)
                        secuencia += 1
                        marcada_inicio = LogMarcada(status=True, logdia=logdia, time=inicio, secuencia=secuencia)
                        marcada_inicio.save()
                        secuencia += 1
                        marcada_fin = LogMarcada(status=True, logdia=logdia, time=fin, secuencia=secuencia)
                        marcada_fin.save()
                    total = logdia.logmarcada_set.all().count()
                    logdia.cantidadmarcadas = total
                    logdia.save()

def migrar_cargos_contratos():
    contratos=PersonaContratos.objects.filter(status=True,denominacionpuesto__isnull=True,cargo__isnull=False)
    total = contratos.count()
    i=0
    cont=0
    for contrato in contratos:
        i+=1
        puesto = DenominacionPuesto.objects.filter(descripcion__unaccent__iexact=contrato.cargo.upper().strip()).first()
        if puesto:
            contrato.denominacionpuesto=puesto
            contrato.save()
            cont+=1
            print(f'{i} de {total}: {puesto.descripcion} & {contrato.cargo}')
        else:
            print(f'{i} de {total}: {contrato.cargo}')
    print(f'Total migrados {cont} de {total}')
migrar_cargos_contratos()

def migrar_unidad_organica_contratos():
    contratos=PersonaContratos.objects.filter(status=True,unidadorganica__isnull=True,unidad__isnull=False)
    total = contratos.count()
    i=0
    cont=0
    for contrato in contratos:
        i+=1
        unidadorganica = Departamento.objects.filter(nombre=contrato.unidad.upper().strip()).first()
        if unidadorganica:
            contrato.unidadorganica=unidadorganica
            contrato.save()
            cont+=1
            print(f'{i} de {total}: {unidadorganica.nombre} & {contrato.unidad}')
        else:
            print(f'{i} de {total}: {contrato.unidad}')
    print(f'Total migrados {cont} de {total}')
migrar_unidad_organica_contratos()

def mover_acciones_tabla_nueva():
    acciones_old=PersonaAcciones.objects.filter(status=True, migrado=False)
    total=acciones_old.count()
    cont = 0
    i = 0
    for accion_old in acciones_old:
        i+=1
        cargo_old = DenominacionPuesto.objects.filter(status=True, descripcion = accion_old.cargo).first()
        unidad_old = Departamento.objects.filter(status=True, nombre = accion_old.unidad).first()
        if cargo_old and unidad_old:
            accion_new= AccionPersonal(persona=accion_old.persona,
                           numerodocumento=accion_old.numerodocumento,
                           tipo=accion_old.tipo,
                           motivoaccion=accion_old.motivo,
                           denominacionpuesto=cargo_old,
                           departamento=unidad_old,
                           rmu=accion_old.remuneracion,
                           explicacion=accion_old.explicacion,
                           fechadesde=accion_old.fecharige,
                           documento=accion_old.ubicacionfisico,
                           estadoarchivo=4,
                           finalizado=True,
                           archivo=accion_old.archivo)
            accion_new.save()
            accion_old.migrado=True
            accion_old.status=False
            accion_old.save()
            cont+=1
            print(f'{cont} de {total} - Migrado {accion_old}: {accion_new}')
        # else:
            # print(f'{i} de {total} - No migrado {accion_old}')
    print(f'total migrados: {cont}')

def comprobar_bienes_constatacion():
    constataciones = ConstatacionFisica.objects.filter(numero__in=[559, 557, 534, 527, 525], periodo_id__isnull=False)
    cont=0
    for constatacion in constataciones:
        cont+=1
        acta = ActaConstatacion.objects.filter(status=True,constatacion=constatacion)
        acta.update(status=False)
        constatacion.estadoacta=1
        constatacion.estado=1
        constatacion.save()

        print('constatacion id:{} - encontrados:{} - total:{} - persona {} - cedula {}'.format(constatacion.id,constatacion.total_encontrados(),
                                                                                   constatacion.t_constataciones(),
                                                                                       constatacion.usuariobienes,constatacion))
    print('Total con errores: {}'.format(cont))

def eliminar_duplicados_actividad_secuencial():
    periodos = PeriodoPlanificacionTH.objects.filter(status=True)
    for periodo in periodos:
        actividadessi = ActividadSecuencialTH.objects.values_list('id',flat=True).filter(status=True,
                                                           producto__gestion__cabecera__periodo=periodo).distinct('actividad')
        actividadesno = ActividadSecuencialTH.objects.values_list('id',flat=True).filter(status=True,
                                                           producto__gestion__cabecera__periodo=periodo).exclude(id__in=actividadessi)
        for actividad in actividadesno:
            print(actividad,",")

# eliminar_duplicados_actividad_secuencial()