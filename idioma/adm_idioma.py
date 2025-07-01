# -*- coding: latin-1 -*-
from datetime import datetime

from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Q
from django.forms import model_to_dict
from django.http import JsonResponse, HttpResponseRedirect, HttpResponse
from django.shortcuts import render

from openpyxl import workbook as openxl
from openpyxl.styles import Font as openxlFont
from openpyxl.styles.alignment import Alignment as alin
from urllib.request import urlopen, Request
import json
from inno.models import *
import random
from decorators import secure_module, last_access
from idioma.forms import PeriodoForm, GrupoForm, PeriodoAsignaturaForm,InscripcionAlumno
from idioma.models import Periodo, Grupo, GrupoInscripcion, PeriodoAsignatura
from settings import EMAIL_DOMAIN
from sga.commonviews import adduserdata
from sga.funciones import MiPaginador, puede_realizar_accion, log, convertir_lista
from sga.models import Empleador, Persona, CUENTAS_CORREOS, RecordAcademico, HistoricoRecordAcademico
from sga.forms import InscripcionCapacitacionForm
from sga.tasks import send_html_mail
from sga.templatetags.sga_extras import encrypt
from django.template.loader import get_template
from sagest.models import Pago
from api.helpers.response_herlper import Helper_Response
from rest_framework import status



@login_required(redirect_field_name='ret', login_url='/loginsga')
#@secure_module
#@last_access
@transaction.atomic()
def view(request):
    data = {}
    adduserdata(request, data)
    data['url_'] = request.path
    persona = request.session['persona']
    periodo = request.session['periodo']

    if request.method == 'POST':
        action = request.POST['action']
        if action == 'addperiodo':
            try:
                f = PeriodoForm(request.POST)
                filter = Q(status=True)
                if f.is_valid():
                    if Periodo.objects.filter(idioma=f.cleaned_data['idioma'],
                                            descripcion=f.cleaned_data['descripcion'].upper().strip(),
                                            fecinicioinscripcion=f.cleaned_data['fecinicioinscripcion'],
                                            fecfininscripcion=f.cleaned_data['fecfininscripcion'],
                                            status=True).exists():
                        return JsonResponse({"result": True, "mensaje": u"El registro ya existe."})
                    periodo = Periodo(idioma=f.cleaned_data['idioma'],
                                  descripcion=f.cleaned_data['descripcion'],
                                  fecinicioinscripcion=f.cleaned_data['fecinicioinscripcion'],
                                  fecfininscripcion=f.cleaned_data['fecfininscripcion'],
                                  estado=f.cleaned_data['estado'])
                    periodo.save(request)

                    log(u'Adicionó nuevo periodo: %s' % periodo, request, "add")
                    return JsonResponse({"result": False, 'mensaje': 'Registro Exitoso'})
                else:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': "bad", "form": [{k: v[0]} for k, v in f.errors.items()],
                                         "mensaje": "Error en el formulario"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})


        elif action == 'editperiodo':
            try:
                f = PeriodoForm(request.POST)
                if f.is_valid():
                    periodo = Periodo.objects.get(pk=encrypt(request.POST['id']))
                    if Periodo.objects.filter(idioma=f.cleaned_data['idioma'],
                                              descripcion=f.cleaned_data['descripcion'].upper().strip(),
                                              fecinicioinscripcion=f.cleaned_data['fecinicioinscripcion'],
                                              fecfininscripcion=f.cleaned_data['fecfininscripcion'],
                                              status=True).exists():
                        return JsonResponse({"result": True, "mensaje": u"El registro ya existe."})
                    # if request.user.has_perm('sga.puede_titulo_tthh'):
                    periodo.idioma = f.cleaned_data['idioma']
                    periodo.descripcion = f.cleaned_data['descripcion']
                    periodo.fecinicioinscripcion = f.cleaned_data['fecinicioinscripcion']
                    periodo.fecfininscripcion = f.cleaned_data['fecfininscripcion']
                    periodo.estado = f.cleaned_data['estado']
                    periodo.save(request)

                    log(u'Modificó periodo idioma: %s' % periodo, request, "edit")
                    return JsonResponse({"result": False, 'mensaje': 'Edicion Exitosa'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": u"Error al guardar los datos."})
        elif action == 'deleteperiodo':
            try:
                periodo = Periodo.objects.get(pk=encrypt(request.POST['id']))
                log(u'Elimino titulo: %s' % periodo, request, "del")
                periodo.status = False
                periodo.save()
                # return JsonResponse({"result": "ok"})
                res_json = {'error': False}
            except Exception as ex:
                transaction.set_rollback(True)
                res_json = {'error': True, "message": "Error: {}".format(ex)}
                # return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})
            return JsonResponse(res_json, safe=False)


        if action == 'addgrupo':
            try:
                f = GrupoForm(request.POST)
                filter = Q(status=True)
                if f.is_valid():
                    if Grupo.objects.filter(periodo_id=int(encrypt(request.GET['id'])),
                                            idcursomoodle=f.cleaned_data['cursomoodle'],
                                            nombre=f.cleaned_data['grupo'].upper().strip(),
                                            cupo=f.cleaned_data['cupo'],
                                            fecinicio=f.cleaned_data['fecinicio'],
                                            horainicio=f.cleaned_data['horainicio'],
                                            fecfin=f.cleaned_data['fecfin'],
                                            horafin=f.cleaned_data['horafin'],
                                            orden=f.cleaned_data['orden'],
                                            status=True).exists():
                        return JsonResponse({"result": True, "mensaje": u"El registro ya existe."})
                    grupo = Grupo(periodo_id=int(encrypt(request.GET['id'])),
                                            idcursomoodle=f.cleaned_data['cursomoodle'],
                                            nombre=f.cleaned_data['grupo'].upper().strip(),
                                            cupo=f.cleaned_data['cupo'],
                                            fecinicio=f.cleaned_data['fecinicio'],
                                            horainicio=f.cleaned_data['horainicio'],
                                            fecfin=f.cleaned_data['fecfin'],
                                            horafin=f.cleaned_data['horafin'],
                                            orden=f.cleaned_data['orden'])
                    grupo.save(request)

                    log(u'Adicionó nuevo grupo: %s' % grupo, request, "add")
                    return JsonResponse({"result": False, 'mensaje': 'Registro Exitoso'})
                else:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': "bad", "form": [{k: v[0]} for k, v in f.errors.items()],
                                         "mensaje": "Error en el formulario"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})


        elif action == 'editgrupo':
            try:
                f = GrupoForm(request.POST)
                if f.is_valid():
                    grupo = Grupo.objects.get(pk=encrypt(request.POST['id']))
                    if Grupo.objects.filter(periodo_id=int(encrypt(request.GET['id'])),
                                            idcursomoodle=f.cleaned_data['cursomoodle'],
                                            nombre=f.cleaned_data['grupo'].upper().strip(),
                                            cupo=f.cleaned_data['cupo'],
                                            fecinicio=f.cleaned_data['fecinicio'],
                                            horainicio=f.cleaned_data['horainicio'],
                                            fecfin=f.cleaned_data['fecfin'],
                                            horafin=f.cleaned_data['horafin'],
                                            orden=f.cleaned_data['orden'],
                                            status=True).exists():
                        return JsonResponse({"result": True, "mensaje": u"El registro ya existe."})
                    # if request.user.has_perm('sga.puede_titulo_tthh'):
                    grupo.idcursomoodle = f.cleaned_data['cursomoodle']
                    grupo.nombre = f.cleaned_data['grupo']
                    grupo.cupo = f.cleaned_data['cupo']
                    grupo.fecinicio = f.cleaned_data['fecinicio']
                    grupo.horainicio = f.cleaned_data['horainicio']
                    grupo.fecfin = f.cleaned_data['fecfin']
                    grupo.horafin = f.cleaned_data['horafin']
                    grupo.orden = f.cleaned_data['orden']
                    grupo.save(request)

                    log(u'Modificó periodo idioma: %s' % grupo, request, "edit")
                    return JsonResponse({"result": False, 'mensaje': 'Edicion Exitosa'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": u"Error al guardar los datos."})

        elif action == 'deletegrupo':
            try:
                grupo = Grupo.objects.get(pk=encrypt(request.POST['id']))
                log(u'Elimino titulo: %s' % grupo, request, "del")
                grupo.status = False
                grupo.save()
                # return JsonResponse({"result": "ok"})
                res_json = {'error': False}
            except Exception as ex:
                transaction.set_rollback(True)
                res_json = {'error': True, "message": "Error: {}".format(ex)}
                # return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})
            return JsonResponse(res_json, safe=False)

        elif action == 'addasignatura':
            try:
                f = PeriodoAsignaturaForm(request.POST)
                filter = Q(status=True)
                if f.is_valid():
                    if PeriodoAsignatura.objects.filter(periodo_id=int(encrypt(request.GET['id'])),
                                                        asignatura=f.cleaned_data['asignatura'],
                                            status=True).exists():
                        return JsonResponse({"result": True, "mensaje": u"El registro ya existe."})
                    periodoasignatura = PeriodoAsignatura(periodo_id=int(encrypt(request.GET['id'])),
                                            asignatura=f.cleaned_data['asignatura'])
                    periodoasignatura.save(request)
                    log(u'Adicionó nueva asignatura al periodo: %s' % periodoasignatura, request, "add")
                    return JsonResponse({"result": False, 'mensaje': 'Registro Exitoso'})
                else:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': "bad", "form": [{k: v[0]} for k, v in f.errors.items()],
                                         "mensaje": "Error en el formulario"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos."})


        elif action == 'editasignatura':
            try:
                f = PeriodoAsignaturaForm(request.POST)
                if f.is_valid():
                    periodoasignatura = PeriodoAsignatura.objects.get(pk=encrypt(request.POST['id']))
                    if PeriodoAsignatura.objects.filter(periodo_id=int(encrypt(request.GET['id'])),
                                                        asignatura=f.cleaned_data['asignatura'],
                                            status=True).exists():
                        return JsonResponse({"result": True, "mensaje": u"El registro ya existe."})
                    # if request.user.has_perm('sga.puede_titulo_tthh'):
                    periodoasignatura.periodo_id = int(encrypt(request.GET['id']))
                    periodoasignatura.asignatura = f.cleaned_data['asignatura']
                    periodoasignatura.save(request)

                    log(u'Modificó asignatura: %s' % periodoasignatura, request, "edit")
                    return JsonResponse({"result": False, 'mensaje': 'Edicion Exitosa'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": u"Error al guardar los datos."})


        elif action == 'deleteasignatura':
            try:
                asignatura = PeriodoAsignatura.objects.get(pk=encrypt(request.POST['id']))
                log(u'Elimino titulo: %s' % asignatura, request, "del")
                asignatura.status = False
                asignatura.save()
                # return JsonResponse({"result": "ok"})
                res_json = {'error': False}
            except Exception as ex:
                transaction.set_rollback(True)
                res_json = {'error': True, "message": "Error: {}".format(ex)}
                # return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})
            return JsonResponse(res_json, safe=False)

        elif action == 'exportarnotasindividual':
            try:
                grupoinscripcion = GrupoInscripcion.objects.get(pk=request.POST['idgrupo'], status=True)
                idcursomoodle = grupoinscripcion.grupo.idcursomoodle
                url = 'https://upei.buckcenter.edu.ec/usernamecoursetograde.php?username=%s&curso=%s' % (grupoinscripcion.inscripcion.persona.identificacion(), idcursomoodle)
                req = Request(url)
                response = urlopen(req)
                result = json.loads(response.read().decode())
                idcurso = int(result['idcurso'])
                #SE VA A UTILIZAR OTRO CAMBPO DEL API EN (NOTA)
                if result['nota'] != 'null':
                    if idcurso == idcursomoodle:
                        if result['nota'] == "-":
                            nota = 0
                        else:
                            nota = null_to_decimal(result['nota'], 0)
                        grupoinscripcion.nota = nota
                        observacion = result['estado']
                        grupoinscripcion.observacion = observacion
                        grupoinscripcion.save()
                    grupoinscripcion.obtener_creditos_horas_modulo()
                    asignatura = None
                    if grupoinscripcion.nota == 70:
                        asignatura = PeriodoAsignatura.objects.filter(periodo_id=int(encrypt(request.POST['id'])),
                                                                      asignatura_id=783, status=True)
                    elif grupoinscripcion.nota > 70 and grupoinscripcion.nota <= 80:
                        asignatura = PeriodoAsignatura.objects.filter(periodo_id=int(encrypt(request.POST['id'])),
                                                                      asignatura_id__in=(783, 784), status=True)
                    elif grupoinscripcion.nota > 80 and grupoinscripcion.nota <= 90:
                        asignatura = PeriodoAsignatura.objects.filter(periodo_id=int(encrypt(request.POST['id'])),
                                                                      asignatura_id__in=(783, 784, 785), status=True)

                    elif grupoinscripcion.nota > 90 and grupoinscripcion.nota <= 100:
                        asignatura = PeriodoAsignatura.objects.filter(periodo_id=int(encrypt(request.POST['id'])),
                                                                      asignatura_id__in=(783, 784, 785, 786),
                                                                      status=True)
                    if asignatura:
                        for asig in asignatura:
                            eMateriaAsignadas=MateriaAsignada.objects.filter(materia__nivel__periodo=periodo,materia__asignatura=asig.asignatura,
                                                                            matricula__inscripcion=grupoinscripcion.inscripcion,
                                                                            materia__nivel_id=1501,matricula__nivel__periodo=periodo)
                            if eMateriaAsignadas.values('id').exists():
                                erubros=eMateriaAsignadas.first().rubro.all()
                                if not Pago.objects.filter(rubro__in=erubros):
                                    erubros.delete()
                                eMateriaAsignadas.delete()

                            if not RecordAcademico.objects.filter(inscripcion=grupoinscripcion.inscripcion, asignatura=asig.asignatura,
                                                                  status=True):
                                notaasignada = 0
                                if grupoinscripcion.nota <= 70:
                                    notaasignada = 70
                                else:
                                    notaasignada = grupoinscripcion.nota
                                record = RecordAcademico(inscripcion=grupoinscripcion.inscripcion,
                                                         asignatura=asig.asignatura,
                                                         grupoinscripcion=grupoinscripcion,
                                                         modulomalla=grupoinscripcion.inscripcion.asignatura_en_modulomalla(
                                                             asig.asignatura),
                                                         nota=notaasignada,
                                                         asistencia=100,
                                                         fecha=datetime.now().date(),
                                                         aprobada=True,
                                                         noaplica=False,
                                                         convalidacion=False,
                                                         pendiente=False,
                                                         creditos=grupoinscripcion.obtener_creditos_horas_modulo()[0][1],
                                                         horas=grupoinscripcion.obtener_creditos_horas_modulo()[0][0],
                                                         homologada=False,
                                                         valida=True,
                                                         validapromedio=False,
                                                         observaciones='Prueba de ubicación de inglés',
                                                         suficiencia=True)
                                record.save(request)
                                if not record.historicorecordacademico_set.filter(status=True, fecha=record.fecha).exists():
                                    nuevohistorico = HistoricoRecordAcademico(recordacademico=record,
                                                                              inscripcion=record.inscripcion,
                                                                              modulomalla=record.modulomalla,
                                                                              asignaturamalla=record.asignaturamalla,
                                                                              asignatura=record.asignatura if record.asignatura else record.asignaturamallahistorico.asignatura,
                                                                              grupoinscripcion=record.grupoinscripcion,
                                                                              nota=record.nota,
                                                                              asistencia=record.asistencia,
                                                                              sinasistencia=record.sinasistencia,
                                                                              fecha=record.fecha,
                                                                              noaplica=record.noaplica,
                                                                              aprobada=record.aprobada,
                                                                              convalidacion=record.convalidacion,
                                                                              homologada=record.homologada,
                                                                              pendiente=record.pendiente,
                                                                              creditos=record.creditos,
                                                                              horas=record.horas,
                                                                              valida=record.valida,
                                                                              validapromedio=record.validapromedio,
                                                                              materiaregular=record.materiaregular,
                                                                              materiacurso=record.materiacurso,
                                                                              observaciones=record.observaciones,
                                                                              completonota=record.completonota,
                                                                              completoasistencia=record.completoasistencia,
                                                                              suficiencia=record.suficiencia)
                                    nuevohistorico.save()
                                seleccionada = record.historicorecordacademico_set.filter(status=True).order_by('-aprobada', '-fecha')[0]
                                record.asignaturamalla = seleccionada.asignaturamalla
                                record.asignatura = seleccionada.asignatura if record.asignatura else None
                                record.grupoinscripcion = seleccionada.grupoinscripcion
                                record.nota = seleccionada.nota
                                record.asistencia = seleccionada.asistencia
                                record.sinasistencia = seleccionada.sinasistencia
                                record.fecha = seleccionada.fecha
                                record.noaplica = seleccionada.noaplica
                                record.aprobada = seleccionada.aprobada
                                record.convalidacion = seleccionada.convalidacion
                                record.homologada = seleccionada.homologada
                                record.pendiente = seleccionada.pendiente
                                record.creditos = seleccionada.creditos
                                record.horas = seleccionada.horas
                                record.valida = seleccionada.valida
                                record.validapromedio = seleccionada.validapromedio
                                record.materiaregular = seleccionada.materiaregular
                                record.materiacurso = seleccionada.materiacurso
                                record.observaciones = seleccionada.observaciones
                                record.completonota = seleccionada.completonota
                                record.completoasistencia = seleccionada.completoasistencia
                                record.suficiencia = seleccionada.suficiencia
                                record.save()
                                grupoinscripcion.inscripcion.actualizar_nivel()
                                grupoinscripcion.inscripcion.actualiza_matriculas(
                                    record.asignatura if record.asignatura else record.asignaturamallahistorico.asignatura)
                                log(u'Adiciono record academico: %s - %s' % (record, record.inscripcion.persona),
                                    request, "add")
                    return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        # elif action == 'migrarmodulosrecordacademico':
        #     try:
        #         grupoinscripcion = GrupoInscripcion.objects.filter(grupo_id=int(encrypt(request.POST['idgrupo'])), status=True).order_by('inscripcion__persona_apellido1')
        #         for list in grupoinscripcion:
        #             idcursomoodle = list.grupo.idcursomoodle
        #             url = 'https://upei.buckcenter.edu.ec/usernamecoursetograde.php?username=%s&curso=%s' % (list.inscripcion.persona.identificacion(), idcursomoodle)
        #             req = Request(url)
        #             response = urlopen(req)
        #             result = json.loads(response.read().decode())
        #             idcurso = int(result['idcurso'])
        #             if result['nota'] != 'null':
        #                 if idcurso == idcursomoodle:
        #                     if result['nota'] == "-":
        #                         nota = 0
        #                     else:
        #                         nota = null_to_decimal(result['nota'], 0)
        #                     list.nota = nota
        #                     list.save()
        #                 list.obtener_creditos_horas_modulo()
        #                 asignatura = None
        #                 if list.nota >= 60 and list.nota <= 70:
        #                     asignatura = PeriodoAsignatura.objects.filter(periodo_id=int(encrypt(request.POST['id'])),
        #                                                                   asignatura_id=783, status=True)
        #                 elif list.nota > 70 and list.nota <= 80:
        #                     asignatura = PeriodoAsignatura.objects.filter(periodo_id=int(encrypt(request.POST['id'])),
        #                                                                   asignatura_id__in=(783, 784), status=True)
        #                 elif list.nota > 80 and list.nota <= 90:
        #                     asignatura = PeriodoAsignatura.objects.filter(periodo_id=int(encrypt(request.POST['id'])),
        #                                                                   asignatura_id__in=(783, 784, 785), status=True)
        #                 elif list.nota > 90 and list.nota <= 100:
        #                     asignatura = PeriodoAsignatura.objects.filter(periodo_id=int(encrypt(request.POST['id'])),
        #                                                                   asignatura_id__in=(783, 784, 785, 786),
        #                                                                   status=True)
        #                 if asignatura:
        #                     for asig in asignatura:
        #                         if not RecordAcademico.objects.filter(inscripcion=list.inscripcion, asignatura=asig.asignatura,
        #                                                               status=True):
        #                             notaasignada = 0
        #                             if list.nota <= 70:
        #                                 notaasignada = 70
        #                             else:
        #                                 notaasignada = list.nota
        #                             record = RecordAcademico(inscripcion=list.inscripcion,
        #                                                      asignatura=asig.asignatura,
        #                                                      grupoinscripcion=list,
        #                                                      modulomalla=list.inscripcion.asignatura_en_modulomalla(
        #                                                          asig.asignatura),
        #                                                      nota=notaasignada,
        #                                                      asistencia=100,
        #                                                      fecha=datetime.now().date(),
        #                                                      aprobada=True,
        #                                                      noaplica=False,
        #                                                      convalidacion=False,
        #                                                      pendiente=False,
        #                                                      creditos=list.obtener_creditos_horas_modulo()[0][1],
        #                                                      horas=list.obtener_creditos_horas_modulo()[0][0],
        #                                                      homologada=False,
        #                                                      valida=True,
        #                                                      validapromedio=False,
        #                                                      observaciones='Prueba de ubicación de inglés',
        #                                                      suficiencia=True)
        #                             record.save(request)
        #                             if not record.historicorecordacademico_set.filter(status=True, fecha=record.fecha).exists():
        #                                 nuevohistorico = HistoricoRecordAcademico(recordacademico=record,
        #                                                                           inscripcion=record.inscripcion,
        #                                                                           modulomalla=record.modulomalla,
        #                                                                           asignaturamalla=record.asignaturamalla,
        #                                                                           asignatura=record.asignatura if record.asignatura else record.asignaturamallahistorico.asignatura,
        #                                                                           grupoinscripcion=record.grupoinscripcion,
        #                                                                           nota=record.nota,
        #                                                                           asistencia=record.asistencia,
        #                                                                           sinasistencia=record.sinasistencia,
        #                                                                           fecha=record.fecha,
        #                                                                           noaplica=record.noaplica,
        #                                                                           aprobada=record.aprobada,
        #                                                                           convalidacion=record.convalidacion,
        #                                                                           homologada=record.homologada,
        #                                                                           pendiente=record.pendiente,
        #                                                                           creditos=record.creditos,
        #                                                                           horas=record.horas,
        #                                                                           valida=record.valida,
        #                                                                           validapromedio=record.validapromedio,
        #                                                                           materiaregular=record.materiaregular,
        #                                                                           materiacurso=record.materiacurso,
        #                                                                           observaciones=record.observaciones,
        #                                                                           completonota=record.completonota,
        #                                                                           completoasistencia=record.completoasistencia,
        #                                                                           suficiencia=record.suficiencia)
        #                                 nuevohistorico.save()
        #                             seleccionada = record.historicorecordacademico_set.filter(status=True).order_by('-aprobada', '-fecha')[0]
        #                             record.asignaturamalla = seleccionada.asignaturamalla
        #                             record.asignatura = seleccionada.asignatura if record.asignatura else None
        #                             record.grupoinscripcion = seleccionada.grupoinscripcion
        #                             record.nota = seleccionada.nota
        #                             record.asistencia = seleccionada.asistencia
        #                             record.sinasistencia = seleccionada.sinasistencia
        #                             record.fecha = seleccionada.fecha
        #                             record.noaplica = seleccionada.noaplica
        #                             record.aprobada = seleccionada.aprobada
        #                             record.convalidacion = seleccionada.convalidacion
        #                             record.homologada = seleccionada.homologada
        #                             record.pendiente = seleccionada.pendiente
        #                             record.creditos = seleccionada.creditos
        #                             record.horas = seleccionada.horas
        #                             record.valida = seleccionada.valida
        #                             record.validapromedio = seleccionada.validapromedio
        #                             record.materiaregular = seleccionada.materiaregular
        #                             record.materiacurso = seleccionada.materiacurso
        #                             record.observaciones = seleccionada.observaciones
        #                             record.completonota = seleccionada.completonota
        #                             record.completoasistencia = seleccionada.completoasistencia
        #                             record.suficiencia = seleccionada.suficiencia
        #                             record.save()
        #                             list.inscripcion.actualizar_nivel()
        #                             list.inscripcion.actualiza_matriculas(
        #                                 record.asignatura if record.asignatura else record.asignaturamallahistorico.asignatura)
        #                             log(u'Adiciono record academico: %s - %s' % (record, record.inscripcion.persona),
        #                                 request, "add")
        #         res_json = {'error': False}
        #     except Exception as ex:
        #         transaction.set_rollback(True)
        #         res_json = {'error': True, "message": "Error: {}".format(ex)}
        #         # return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})
        #     return JsonResponse(res_json, safe=False)

        elif action == 'traeralumnos':
            try:
                listaenviar = GrupoInscripcion.objects.filter(grupo_id=int(encrypt(request.POST['idgrupo'])), status=True).values('id', 'inscripcion__persona__apellido1', 'inscripcion__persona__apellido2', 'inscripcion__persona__nombres').order_by('inscripcion__persona_apellido1')
                return JsonResponse({"result": "ok", "cantidad": len(listaenviar), "inscritos": convertir_lista(listaenviar)})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al listar los estudiantes."})

        elif action == 'deleteinscip':
            try:
                idioma = GrupoInscripcion.objects.get(inscripcion=request.POST['id'],grupo=request.POST['grupo'])
                if idioma.recordacademico_set.exists():
                    return JsonResponse({'result': 'bad', 'mensaje': u'No puede eliminar porque esta inscripcion porque este modulo ya consta en el record academico'})
                log(u"Elimino a al alumno : %s" % idioma, request, "del")
                idioma.delete()
                return JsonResponse({'result': 'ok'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': 'bad', 'mensaje': u'Error al eliminar los datos'})


        if action == 'registraralumno':
            try:
                eRequest = InscripcionAlumno(request.POST)
                ESTADO_SOLICITADO = 0
                eGrupo = Grupo.objects.get(pk=int(encrypt(request.POST['idgrupo'])))
                if eRequest.is_valid():
                    if GrupoInscripcion.objects.values('id').filter(status=True, grupo=eGrupo,
                                                                    inscripcion=eRequest.cleaned_data['inscripcion']).exists():
                        raise NameError(u'Ya se encuentra registrada su inscripción')
                    if not eGrupo.existe_cupo_disponible():
                        raise NameError(
                            u'Lo sentimos, no existe cupos disponibles en este grupo, actualice e intente de nuevo.')
                    eGrupoInscripcion = GrupoInscripcion(grupo=eGrupo, inscripcion=eRequest.cleaned_data['inscripcion'], estado=ESTADO_SOLICITADO)
                    eGrupoInscripcion.save(request)
                    aData = {}
                    log(u'Adicionó nueva alumno al grupo: %s' % eGrupoInscripcion, request, "add")
                    return JsonResponse({"result": False, 'mensaje': 'Registro Exitoso'})
                else:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in eRequest.errors.items()],
                                         "mensaje": "Error en el formulario"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": f'Ocurrio un error: {ex.__str__()}'})

        # elif action == 'importarnotaexamen':
        #     try:
        #         grupoinscripcion = GrupoInscripcion.objects.filter(grupo_id=int(encrypt(request.POST['idgrupo'])), status=True)
        #         for list in grupoinscripcion:
        #             idcursomoodle = list.grupo.idcursomoodle
        #             url = 'https://upei.buckcenter.edu.ec/usernamecoursetograde.php?username=0942301706&curso=96'
        #             #% (list.inscripcion.persona.identificacion(), idcursomoodle)
        #             req = Request(url)
        #             response = urlopen(req)
        #             result = json.loads(response.read().decode())
        #             idcurso = int(result['idcurso'])
        #             if idcurso != idcursomoodle:
        #                 nota = null_to_decimal(result['nota'], 0)
        #                 if nota == '-':
        #                     nota = 0
        #                 list.nota = nota
        #                 list.save()
        #             #return JsonResponse({"result": "ok"})
        #         res_json = {'error': False}
        #     except Exception as ex:
        #         transaction.set_rollback(True)
        #         res_json = {'error': True, "message": "Error: {}".format(ex)}
        #         # return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})
        #     return JsonResponse(res_json, safe=False)


        return JsonResponse({"result": False, "mensaje": u"Solicitud Incorrecta."})
    else:
        if 'action' in request.GET:
            data['action'] = action = request.GET['action']
            if action == 'addperiodo':
                try:
                    data['title'] = u'Adicionar nuevo Periodo'
                    data['action'] = request.GET['action']
                    data['form'] = PeriodoForm()
                    template = get_template("idioma/modal/formPeriodo.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, 'message': str(ex)})
            if action == 'editperiodo':
                try:
                    data['title'] = u'Modificar Periodo'
                    data['action'] = request.GET['action']
                    data['id'] = id = int(encrypt(request.GET['id']))
                    data['periodo'] = periodo = Periodo.objects.get(pk=int(encrypt(request.GET['id'])))
                    form = PeriodoForm(initial={'idioma': periodo.idioma,
                                              'descripcion': periodo.descripcion,
                                              'fecinicioinscripcion': periodo.fecinicioinscripcion.date(),
                                              'fecfininscripcion': periodo.fecfininscripcion.date(),
                                              'estado': periodo.estado})
                    data['form'] = form
                    template = get_template("idioma/modal/formPeriodo.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, 'message': str(ex)})

            if action == 'vergrupos':
                try:
                    data['title'] = u'Gestión de grupos'
                    data['id'] = id = int(encrypt(request.GET['id']))
                    data['periodo'] = nomperiodo = Periodo.objects.get(pk=id,status=True)
                    search, filtro, url_vars = request.GET.get('s', ''), (Q(status=True, periodo_id=id)), f'&action={action}&id={request.GET["id"]}'
                    if search:
                        data['search'] = search
                        url_vars += "&s={}".format(search)
                        filtro = filtro & (Q(nombre__icontains=search))
                    grupos = Grupo.objects.filter(filtro).order_by('fecha_creacion')
                    paging = MiPaginador(grupos, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['listado'] = page.object_list
                    data["url_vars"] = url_vars
                    data['list_count'] = len(grupos)
                    return render(request, "idioma/vergrupos.html", data)
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, 'message': str(ex)})

            if action == 'addgrupo':
                try:
                    data['title'] = u'Adicionar nuevo Grupo'
                    data['action'] = request.GET['action']
                    data['form'] = GrupoForm()
                    template = get_template("idioma/modal/formGrupo.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, 'message': str(ex)})

            if action == 'editgrupo':
                try:
                    data['title'] = u'Modificar grupo'
                    data['action'] = request.GET['action']
                    data['id'] = id = int(encrypt(request.GET['id']))
                    data['grupo'] = grupo = Grupo.objects.get(pk=int(encrypt(request.GET['id'])))
                    form = GrupoForm(initial={'grupo': grupo.nombre,
                                              'cursomoodle': grupo.idcursomoodle,
                                              'fecinicio': grupo.fecinicio,
                                              'horainicio': grupo.horainicio,
                                              'fecfin': grupo.fecfin,
                                              'horafin': grupo.horafin,
                                              'cupo': grupo.cupo,
                                              'orden': grupo.orden})
                    data['form'] = form
                    template = get_template("idioma/modal/formGrupo.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, 'message': str(ex)})

            if action == 'verinscritos':
                try:
                    data['title'] = u'Listado de inscritos'
                    data['id'] = id = int(encrypt(request.GET['idg']))
                    search, filtro, url_vars = request.GET.get('s', ''), (
                        Q(grupo_id=id, status=True)), f'&action={action}&idg={request.GET["idg"]}'
                    if search:
                        data['search'] = search
                        url_vars += "&s={}".format(search)
                        filtro = filtro & (Q(inscripcion__persona__nombres__icontains=search) |
                                           Q(inscripcion__persona__apellido1__icontains=search) |
                                           Q(inscripcion__persona__apellido2__icontains=search)|
                                           Q(inscripcion__persona__cedula__icontains=search))
                    inscritos = GrupoInscripcion.objects.filter(filtro).order_by('inscripcion__persona_apellido1')
                    data['grupo'] = Grupo.objects.get(pk=id)
                    paging = MiPaginador(inscritos, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['listado'] = page.object_list
                    data["url_vars"] = url_vars
                    data['list_count'] = len(inscritos)
                    return render(request, "idioma/verinscritos.html", data)
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, 'message': str(ex)})

            if action == 'verasignaturas':
                try:
                    data['title'] = u'Gestión de asignaturas'
                    data['id'] = id = int(encrypt(request.GET['id']))
                    data['periodo'] = nomperiodo = Periodo.objects.get(pk=id,status=True)
                    search, filtro, url_vars = request.GET.get('s', ''), (Q(status=True, periodo_id=id)), f'&action={action}&id={request.GET["id"]}'
                    # if search:
                    #     data['search'] = search
                    #     url_vars += "&s={}".format(search)
                    #     filtro = filtro & (Q(nombre__icontains=search))
                    asignaturas = PeriodoAsignatura.objects.filter(filtro).order_by('fecha_creacion')
                    paging = MiPaginador(asignaturas, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['listado'] = page.object_list
                    data["url_vars"] = url_vars
                    data['list_count'] = len(asignaturas)
                    return render(request, "idioma/verasignaturas.html", data)
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, 'message': str(ex)})


            if action == 'addasignatura':
                try:
                    data['title'] = u'Adicionar nueva Asignatura'
                    data['action'] = request.GET['action']
                    data['form'] = PeriodoAsignaturaForm()
                    template = get_template("idioma/modal/formPeriodoAsignatura.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, 'message': str(ex)})

            if action == 'editasignatura':
                try:
                    data['title'] = u'Modificar Asignatura'
                    data['action'] = request.GET['action']
                    data['id'] = id = int(encrypt(request.GET['id']))
                    data['asignatura'] = asignatura = PeriodoAsignatura.objects.get(pk=int(encrypt(request.GET['id'])))
                    form = PeriodoAsignaturaForm(initial={'asignatura': asignatura.asignatura})
                    data['form'] = form
                    template = get_template("idioma/modal/formPeriodoAsignatura.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, 'message': str(ex)})

            if action == 'registraralumno':
                try:
                    data['title'] = u'Modificar Asignatura'
                    data['action'] = request.GET['action']
                    data['idperiodo'] = idper = int(encrypt(request.GET['idperiodo']))
                    data['idgrupo'] = idgru = int(encrypt(request.GET['idgrupo']))
                    data['asignatura'] = asignatura = PeriodoAsignatura.objects.get(
                        pk=idper)
                    form = InscripcionAlumno()
                    form.fields['inscripcion'].queryset = Inscripcion.objects.none()
                    data['form'] = form
                    template = get_template("idioma/modal/formRegistro.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, 'message': str(ex)})

            if action == 'buscarpersona':
                try:
                    q = request.GET['q'].upper().strip()
                    s = q.split(" ")
                    if len(s) == 1:
                        per = Inscripcion.objects.filter(
                            (Q(persona__nombres__icontains=q) | Q(persona__apellido1__icontains=q) | Q(
                                persona__apellido2__icontains=q) | Q(persona__cedula__contains=q)),
                            Q(status=True, activo=True),matricula__nivel__periodo=periodo)[:15]
                    elif len(s) == 2:
                        per = Inscripcion.objects.filter(
                            (Q(persona__apellido1__icontains=s[0]) & Q(persona__apellido2__icontains=s[1])) |
                            (Q(persona__nombres__icontains=s[0]) & Q(persona__nombres__icontains=s[1])) |
                            (Q(persona__nombres__icontains=s[0]) & Q(persona__apellido1__icontains=s[1]))).filter(
                            Q(status=True, activo=True),matricula__nivel__periodo=periodo)[:15]
                    data = [{'id': qs.pk, 'text': f"{qs.persona.nombre_completo_inverso()}",
                                 'documento': qs.persona.documento(), 'carrera': qs.carrera.nombre,
                                 'foto': qs.persona.get_foto()} for qs in per]
                    return HttpResponse(json.dumps({'status': True, 'results': data}))
                except Exception as ex:
                    pass

            return HttpResponseRedirect(request.path)
        else:
            try:
                data['title'] = u'Gestión de periodos'
                search, filtro, url_vars = request.GET.get('s', ''), (Q(status=True)), ''
                if search:
                    data['search'] = search
                    url_vars += "&s={}".format(search)
                    filtro = filtro & (Q(descripcion__icontains=search))
                if 'id' in request.GET:
                    ids = int(request.GET['id'])
                    url_vars += "&id={}".format(ids)
                    filtro = filtro & Q(pk=ids)
                periodo= Periodo.objects.filter(filtro).order_by('fecha_creacion')
                paging = MiPaginador(periodo, 20)
                p = 1
                try:
                    paginasesion = 1
                    if 'paginador' in request.session:
                        paginasesion = int(request.session['paginador'])
                    if 'page' in request.GET:
                        p = int(request.GET['page'])
                    else:
                        p = paginasesion
                    try:
                        page = paging.page(p)
                    except:
                        p = 1
                    page = paging.page(p)
                except:
                    page = paging.page(p)
                request.session['paginador'] = p
                data['paging'] = paging
                data['rangospaging'] = paging.rangos_paginado(p)
                data['page'] = page
                data['listado'] = page.object_list
                data["url_vars"] = url_vars
                data['list_count'] = len(periodo)

                # EXPORTAR EXCEL PARA VISUALIZAR INSCRITOS
                if 'exportar_excel' in request.GET:
                    periodo = int(encrypt(request.GET['periodo']))
                    inscritos = GrupoInscripcion.objects.filter(status=True, grupo__periodo_id=periodo).order_by(
                        'grupo__nombre', 'inscripcion__persona__apellido1')
                    wb = openxl.Workbook()
                    wb["Sheet"].title = "Reporte_inscritos"
                    ws = wb.active
                    style_title = openxlFont(name='Arial', size=16, bold=True)
                    style_cab = openxlFont(name='Arial', size=10, bold=True)
                    alinear = alin(horizontal="center", vertical="center")
                    response = HttpResponse(content_type="application/ms-excel")
                    response[
                        'Content-Disposition'] = 'attachment; filename=Reporte de inscritos' + '-' + random.randint(
                        1, 10000).__str__() + '.xlsx'
                    ws.column_dimensions['B'].width = 15
                    ws.column_dimensions['C'].width = 15
                    ws.column_dimensions['D'].width = 40
                    ws.column_dimensions['E'].width = 35
                    ws.column_dimensions['F'].width = 55
                    ws.column_dimensions['G'].width = 60
                    ws.column_dimensions['H'].width = 25
                    ws.column_dimensions['I'].width = 15
                    ws.column_dimensions['J'].width = 15
                    ws.column_dimensions['K'].width = 15
                    ws.column_dimensions['L'].width = 20
                    ws.column_dimensions['M'].width = 15
                    ws.column_dimensions['N'].width = 15
                    ws.column_dimensions['O'].width = 15
                    ws.merge_cells('A1:O1')
                    ws['A1'] = 'LISTADO DE INSCRITOS (PRUEBA DE INGLES)'
                    celda1 = ws['A1']
                    celda1.font = style_title
                    celda1.alignment = alinear

                    columns = [u"N°", u"CEDULA",u"INSCRIPCION",u"ESTUDIANTE", u"CORREO",
                               u"FACULTAD",u"CARRERA", u"CONVOCATORIA", u"GRUPO", u"ID_GRUPO",
                               u"FECHA", u"HORA",u"ESTADO", u"ID_MOODLE", u"NOTA",u"OBSERVACIÓN",
                               ]
                    row_num = 2
                    for col_num in range(0, len(columns)):
                        celda = ws.cell(row=row_num, column=(col_num + 1), value=columns[col_num])
                        celda.font = style_cab
                    row_num = 3
                    numero = 1
                    for list in inscritos:
                        ws.cell(row=row_num, column=1, value=numero)
                        ws.cell(row=row_num, column=2, value=str(list.inscripcion.persona.cedula))
                        ws.cell(row=row_num, column=3, value=str(list.inscripcion.id))
                        ws.cell(row=row_num, column=4, value=str(list.inscripcion.persona.nombre_completo_minus()))
                        ws.cell(row=row_num, column=5,
                                value=str(list.inscripcion.persona.emailinst) if list.inscripcion.persona.emailinst else list.inscripcion.persona.email)
                        ws.cell(row=row_num, column=6, value=str(list.inscripcion.coordinacion.nombre))
                        ws.cell(row=row_num, column=7, value=str(list.inscripcion.carrera.nombre))
                        ws.cell(row=row_num, column=8, value=str(list.grupo.periodo.descripcion))
                        ws.cell(row=row_num, column=9, value=str(list.grupo.nombre))
                        ws.cell(row=row_num, column=10, value=str(list.grupo.id))
                        ws.cell(row=row_num, column=11, value=str(list.grupo.fecinicio))
                        ws.cell(row=row_num, column=12, value=str(list.grupo.horainicio)  + ' - ' +  str(list.grupo.horafin))
                        if list.estado == 0:
                            estado = 'En curso'
                        elif list.estado == 1:
                            estado = 'Aprobado'
                        else:
                            estado = 'Reprobado'
                        ws.cell(row=row_num, column=13, value=str(estado))
                        ws.cell(row=row_num, column=14, value=str(list.grupo.idcursomoodle))
                        ws.cell(row=row_num, column=15, value=str(list.nota))
                        ws.cell(row=row_num, column=16, value=str(list.observacion))
                        row_num += 1
                        numero += 1
                    wb.save(response)
                    return response

                return render(request, "idioma/view.html", data)
            except Exception as ex:
                pass
