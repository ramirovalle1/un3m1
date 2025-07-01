# -*- coding: UTF-8 -*-
import random
import sys
import calendar
from datetime import date

from openpyxl import workbook as openxl
from openpyxl.chart import ScatterChart, Reference, Series,PieChart, BarChart
from openpyxl.styles import Font as openxlFont
from openpyxl.styles.alignment import Alignment as alin

from xlwt import *

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Q
import xlwt
from django.forms import model_to_dict
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.template import Context
from django.template.loader import get_template
from django.shortcuts import render, redirect
from decorators import secure_module
from settings import EMAIL_DOMAIN
from sga.commonviews import adduserdata
from sga.templatetags.sga_extras import encrypt
from sga.funciones import MiPaginador, log, generar_nombre, remover_caracteres_especiales_unicode
from sga.models import Administrativo, Persona, Utensilios, MESES_CHOICES, PersonaDatosFamiliares
from poli.forms import FinalizaReservaForm
from poli.models import *
from django.db.models import Sum, Q, F, FloatField


@login_required(redirect_field_name='ret', login_url='/loginsga')
@secure_module
@transaction.atomic()
def view(request):
    data = {}
    adduserdata(request, data)
    hoy = datetime.now().date()
    usuario = request.user
    persona = request.session['persona']
    periodo = request.session['periodo']
    perfilprincipal = request.session['perfilprincipal']
    data['es_instructor']=es_instructor=persona.instructorpolideportivo_set.filter(status=True).exists()
    if request.method == 'POST':
        res_json = []
        action = request.POST['action']

        if action == 'addimplemento':
            try:
                with transaction.atomic():
                    reservacion = ReservacionPersonaPoli.objects.get(pk=int(request.POST['id']))
                    implementos = request.POST.getlist('infoImplemento[]')
                    if implementos:
                        c = 0
                        while c < len(implementos):
                            implemento = ImplementosReservacionPersona(reservacion=reservacion,
                                                                  utensilio_id=int(implementos[c]),
                                                                  cantidad=int(implementos[c + 1]))
                            implemento.save(request)
                            c += 2
                        implemento.save(request)
                        log(u'Adicionó implemento: %s' % implemento, request, "add")
                        return JsonResponse({"result": False}, safe=False)
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({"result": True, "mensaje": "Registre al menos un implemento."},
                                            safe=False)

            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        if action == 'finalizar':
            try:
                with transaction.atomic():
                    instance = ReservacionPersonaPoli.objects.get(pk=int(request.POST['id']))
                    form = FinalizaReservaForm(request.POST)
                    if form.is_valid():
                        instance.estado = form.cleaned_data['estado']
                        instance.observacion = form.cleaned_data['observacion']
                        instance.asistio=form.cleaned_data['asistio']
                        instance.save(request)
                        if instance.estado == 3:
                            reservafechas = instance.reservacionfechaspoli_set.get(status=True)
                            reservafechas.status = False
                            reservafechas.save(request)

                            reservaturno = reservafechas.reservacionturnospoli_set.get(status=True)
                            reservaturno.status = False
                            reservaturno.save(request)

                            reservafamiliar = reservaturno.reservaciontercerospoli_set.filter(status=True)
                            if reservafamiliar:
                                for familiar in reservafamiliar:
                                    familiar.status = True
                                    familiar.save(request)

                            reservacionseccion = reservaturno.reservacionseccionpoli_set.filter(status=True)
                            for reservaseccion in reservacionseccion:
                                reservaseccion.status = False
                                reservaseccion.save(request)
                        log(u'Cambió estado a reserva: %s' % instance, request, "add")
                        return JsonResponse({"result": False}, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        if action == 'delreserva':
            try:
                with transaction.atomic():
                    reserva = ReservacionPersonaPoli.objects.get(id=int(encrypt(request.POST['id'])))
                    reserva.status = False
                    reserva.save(request)
                    if reserva.estado != 3:
                        reservafechas=reserva.reservacionfechaspoli_set.get(status=True)
                        reservafechas.status=False
                        reservafechas.save(request)

                        reservaturno = reservafechas.reservacionturnospoli_set.get(status=True)
                        reservaturno.status = False
                        reservaturno.save(request)

                        reservafamiliar=reservaturno.reservaciontercerospoli_set.filter(status=True)
                        if reservafamiliar:
                            for familiar in reservafamiliar:
                                familiar.status=True
                                familiar.save(request)

                        reservacionseccion = reservaturno.reservacionseccionpoli_set.filter(status=True)
                        for reservaseccion in reservacionseccion:
                            reservaseccion.status = False
                            reservaseccion.save(request)
                    log(u'Elimino Reservacion: %s' % reserva, request, "delreserva")
                    res_json = {"error": False}
            except Exception as ex:
                res_json = {'error': True, "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)

        if action == 'buscarpersona':
            try:
                item = []
                param = request.POST['term']
                filtros=Q(status=True)
                if param:
                    data['search'] = param
                    s = param.split()
                    if len(s) == 1:
                        filtros = filtros & (Q(apellido2__icontains=param) | Q(cedula__icontains=param) |
                                             Q(apellido1__icontains=param) |Q(apellido2__icontains=param) |
                                             Q(cedula__icontains=param) | Q(apellido1__icontains=param))
                    else:
                        filtros = filtros & (Q(apellido1__icontains=s[0]) & Q(apellido2__icontains=s[1])
                                             | Q(apellido1__icontains=s[0]) & Q(apellido2__icontains=s[1]))

                personas = Persona.objects.filter(filtros).order_by('nombres')
                for persona in personas:
                    text = persona.__str__()
                    item.append({'id': persona.id, 'text': text})
                return JsonResponse(item, safe=False)
            except Exception as e:
                pass

        if action == 'actividades':
            try:
                lista = []
                idarea = int(request.POST['idarea'])
                idperfil=int(request.POST['idperfil'])
                perfilusuario = PerfilUsuario.objects.get(status=True, id=idperfil)
                actividades = ActividadPolideportivo.objects.filter(status=True,mostrar=True, area_id=idarea)
                perfil = None
                carrera = None
                if perfilusuario.es_estudiante():
                    perfil = 1
                    carrera = perfilusuario.inscripcion.carrera
                    actividades = actividades.filter(Q(carreras=carrera) | Q(carreras=None))
                elif perfilusuario.es_administrativo():
                    perfil = 2
                elif perfilusuario.es_profesor():
                    perfil = 3
                elif perfilusuario.es_externo():
                    perfil = 4
                for a in actividades:
                    perfiles = PerfilesActividad.objects.filter(status=True, actividad=a)
                    if perfiles.filter(perfil=perfil, activo=True).exists():
                        text = str(a)
                        lista.append([a.id, text])
                return JsonResponse({'result': 'ok', 'lista': lista})
            except Exception as e:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        if action == 'perfiles':
            try:
                lista = []
                idpersona = int(request.POST['idpersona'])
                perfiles=PerfilUsuario.objects.filter(status=True, persona_id=idpersona)
                for p in perfiles:
                    text = str(p)
                    if p.inscripcion:
                        text = "{} ({})".format(str(p), str(p.inscripcion.carrera))
                    lista.append([p.id, text])
                return JsonResponse({'result': 'ok', 'lista': lista})
            except Exception as e:
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        if action == 'reservar':
            try:
                fecha=datetime.strptime(request.POST['fecha'],'%Y-%m-%d').date()
                cantidad=0
                tercero=False
                tipotercero=None
                if 'tercero' in request.POST:
                    tercero=True
                    # if not 'tipotercero' in request.POST:
                    #     return JsonResponse(
                    #         {"result": True, "mensaje": "Seleccione el tipo de terceros que desea agregar."})
                    tipotercero= 3
                if not request.POST['horario']:
                    transaction.set_rollback(True)
                    return JsonResponse(
                        {"result": True, "mensaje": "Seleccione un horario a reservar."})
                perfil=PerfilUsuario.objects.get(pk=int(request.POST['idperfil']))
                horario=HorarioActividadPolideportivo.objects.get(id=request.POST['horario'])
                seccionesarea = horario.actividad.area.secciones()
                turno=horario.generar_turno(str(perfil.persona))
                reserva = ReservacionPersonaPoli(persona=perfil.persona,
                                                 area=horario.actividad.area,
                                                 estado= 2,
                                                 actividad=horario.actividad,
                                                 tercero=tercero,
                                                 tipotercero=tipotercero,
                                                 codigo=turno)
                reserva.perfil = perfil
                if perfil.persona.es_estudiante():
                    reserva.inscripcion = perfil.inscripcion
                reserva.finicialreserva = fecha
                reserva.ffinalreserva = fecha

                turnos = horario.cupos_reservados(fecha)
                reserva.save(request)

                reservacionfecha=ReservacionFechasPoli(reservacion=reserva, freservacion=hoy)
                reservacionfecha.save(request)

                reservacionturno = ReservacionTurnosPoli(fechareservacion=reservacionfecha, turno=horario)
                reservacionturno.save(request)

                if tercero:
                    grupointerno = request.POST.getlist('grupointerno[]')
                    cantidad = len(grupointerno)
                    maxgrupo = horario.actividad.numacompanantes
                    if cantidad > maxgrupo:
                        transaction.set_rollback(True)
                        return JsonResponse(
                            {"result": True, "mensaje": "Cantida máxima de acompañantes permitido {}".format(maxgrupo)})
                    if grupointerno:
                        f = 0
                        if not cantidad <= turnos:
                            transaction.set_rollback(True)
                            return JsonResponse({"result": True, "mensaje": "La cantidad de acompañantes excede el limite por turno."})
                        while f < len(grupointerno):
                            reservatercero = ReservacionTercerosPoli(reservacion=reservacionturno, persona_id=int(grupointerno[f]))
                            reservatercero.save(request)
                            f += 1
                        reservatercero.save(request)
                        log(u'Adicionó Terceros: %s' % reservatercero, request, "add")
                        # return JsonResponse({"result": False}, safe=False)
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({"result": True, "mensaje": "Registre al menos un acompañante."})

                if seccionesarea:
                    if not 'seccion[]' in request.POST:
                        transaction.set_rollback(True)
                        return JsonResponse(
                            {"result": True, "mensaje": 'Por favor seleccione la sección en la que desea reservar'})
                    cantidad+=1
                    seccionesmarcadas = request.POST.getlist('seccion[]')
                    if cantidad > 1:
                        cupos = horario.actividad.area.cantidad_cupo_secciones(horario,seccionesmarcadas, fecha)
                        if cantidad <= cupos:
                            comprobador=cantidad
                            cont=0
                            cupos1=0
                            listaordenada=[]
                            dicordenado=horario.actividad.area.listado_cupos(horario,seccionesmarcadas, fecha)
                            for dic in dicordenado:
                                cont+=1
                                cupos1+=dic.get('cupos')
                                if comprobador <= cupos1 and len(seccionesmarcadas) > 1 and cont < len(seccionesmarcadas):
                                    transaction.set_rollback(True)
                                    return JsonResponse(
                                        {"result": True,
                                         "mensaje": 'La cantidad de externos o familiares pueden ingresar en menos secciones que las que selecciono.'})
                                listaordenada.append(dic.get('id'))
                            for idseccion in listaordenada:
                                seccion=seccionesarea.get(id=idseccion)
                                seccionreserva = ReservacionSeccionPoli(reservacionturno=reservacionturno,seccion=seccion)
                                seccionreserva.save(request)
                                reservacionturno.seccion.add(seccion)
                                cuposeccion = seccion.cupo_disponible_seccion(fecha,horario)
                                # if tipotercero == 2:
                                if cantidad >= cuposeccion:
                                    cantidad-=cuposeccion
                                    seccionreserva.cantidad=cuposeccion
                                elif cantidad != 0:
                                    seccionreserva.cantidad=cantidad
                                    cantidad = 0
                                seccionreserva.save(request)
                        else:
                            transaction.set_rollback(True)
                            return JsonResponse(
                                {"result": True, "mensaje": 'El número de secciones seleccionados no cubre el número de extras asignados'})
                    elif len(seccionesmarcadas) == 1:
                        idseccion=eval(seccionesmarcadas[0])
                        seccionreserva = ReservacionSeccionPoli(reservacionturno=reservacionturno, seccion_id=idseccion, cantidad=cantidad)
                        seccionreserva.save(request)
                        for seccion in seccionesmarcadas:
                            reservacionturno.seccion.add(seccion)
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse(
                            {"result": True, "mensaje": 'Al ser solo una reservación para su persona, solo debe seleccionar una sección.'})
                log(u'{} : Inicio Reservacion de Area Polideportivo {}'.format(perfil.persona, horario.actividad.area.__str__()), request, "add")
                url_ = '{}?action=admreservar'.format(request.path)
                return JsonResponse({"result": False, "to": url_})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al obtener los datos."})

        if action == 'editimplemento':
            try:
                implemento = ImplementosActividad.objects.get(id=int(encrypt(request.POST['id'])))
                if type(eval(request.POST['valor'])) is int:
                    cantidad = float(request.POST['valor'])
                    if cantidad <= 0:
                        return JsonResponse({"result": False, 'mensaje': 'No puede ingresar una cantidad menor o igual a 0'},
                                            safe=False)
                    implemento.cantidad = cantidad
                    mensaje='Cantidad Actualizada'
                else:
                    implemento.activo = eval(request.POST['valor'])
                    mensaje = 'Uso Actualizado'
                implemento.save(request)
                log('Edito Cantidad de implementos %s' % implemento,request,'editimplemento',)
                return JsonResponse({"result": True, 'mensaje':mensaje}, safe=False)
            except Exception as e:
                return JsonResponse({"result": False, 'mensaje': str(e)}, safe=False)

        if action == 'editasistenciatercero':
            try:
                tercero = ReservacionTercerosPoli.objects.get(id=int(encrypt(request.POST['id'])))
                tercero.asistio = eval(request.POST['valor'])
                mensaje = 'Asistencia de Familiar Actualizado'
                tercero.save(request)
                log('Edito Asisitencia de familiar %s' % tercero,request,'editasistenciatercero',)
                return JsonResponse({"result": True, 'mensaje':mensaje}, safe=False)
            except Exception as e:
                return JsonResponse({"result": False, 'mensaje': str(e)}, safe=False)

        if action =='addobservacion':
            try:
                reservacion=int(encrypt(request.POST['id']))
                observaciones = request.POST['observaciones'].split(',')
                for observacion in observaciones:
                    observacion = ObservacionReservacionPersona(reservacion_id=reservacion, observacion=observacion)
                    observacion.save(request)
                mensaje='Observaciones Agregadas'
                log('Adiciono Nueva Observacion %s' % observacion, request, 'addobservacion')
                return JsonResponse({"result": True, 'mensaje': mensaje}, safe=False)
            except Exception as e:
                return JsonResponse({"result": False, 'mensaje': str(e)}, safe=False)

        if action == 'editobservacion':
            try:
                observacionpersona = ObservacionReservacionPersona.objects.get(id=int(encrypt(request.POST['id'])))
                observacionpersona.observacion = request.POST['valor']
                observacionpersona.save(request)
                mensaje = 'Observación Actualizada'
                log('Edito Cantidad de implementos %s' % observacionpersona, request, 'editobservacion', )
                return JsonResponse({"result": True, 'mensaje': mensaje}, safe=False)
            except Exception as e:
                return JsonResponse({"result": False, 'mensaje': str(e)}, safe=False)

        if action == 'delobservacion':
            try:
                with transaction.atomic():
                    observacion = ObservacionReservacionPersona.objects.get(id=int(encrypt(request.POST['id'])))
                    observacion.status=False
                    observacion.save(request)
                    log(u'Elimino Observacion de reserva: %s' % observacion, request, "delobservacion")
                    res_json = {"error": False}
            except Exception as ex:
                res_json = {'error': True, "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)

        if action == 'buscarpersonas':
            try:
                item = []
                param = request.POST['term'].upper().strip()
                idpersona=int(request.POST['idpersona'])
                s = param.split(" ")
                perfiles=PerfilUsuario.objects.filter((Q(administrativo__isnull=False) | Q(inscripcion__isnull=False) | Q(profesor__isnull=False)), status=True, externo__isnull=True).order_by('persona_id').distinct('persona_id').values_list('persona_id')
                personas=Persona.objects.filter(status=True, id__in=perfiles).exclude(id=idpersona)
                if len(s) == 1:
                    personas = personas.filter((Q(nombres__icontains=param) |
                                           Q(apellido1__icontains=param) |
                                           Q(apellido2__icontains=param) |
                                           Q(cedula__contains=param))).distinct()[:15]
                elif len(s) == 2:
                    personas = personas.filter((Q(apellido1__contains=s[0]) & Q(apellido2__contains=s[1])) |
                                          (Q(nombres__icontains=s[0]) & Q(nombres__icontains=s[1])) |
                                          (Q(nombres__icontains=s[0]) & Q(apellido1__contains=s[1]))).distinct()[:15]
                else:
                    personas = personas.filter((Q(nombres__contains=s[0]) & Q(apellido1__contains=s[1]) & Q(apellido2__contains=s[2])) |
                                               (Q(nombres__contains=s[0]) & Q(nombres__contains=s[1]) & Q(apellido1__contains=s[2]))).distinct()[:15]

                for p in personas:
                    text = str(p)
                    item.append({'id': p.id, 'text': text})

                return JsonResponse(item, safe=False)
            except Exception as e:
                pass

        return JsonResponse({"result": "bad", "mensaje": u"Solicitud Incorrecta."})

    else:
        if 'action' in request.GET:
            data['action'] = action = request.GET['action']

            if action == 'addimplemento':
                try:
                    data['reservacion'] = reservacion = ReservacionPersonaPoli.objects.get(pk=request.GET['id'])
                    data['id'] = request.GET['id']
                    data['detalles'] = reservacion.implementosreservacionpersona_set.filter(status=True)
                    id_excluir = reservacion.implementosreservacionpersona_set.values_list('id',flat=True).filter(status=True)
                    data['implementos'] = Utensilios.objects.filter(status=True).exclude(pk__in=id_excluir)
                    template = get_template("adm_revisareserva/modal/implementos.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'finalizar':
                try:
                    form = FinalizaReservaForm()
                    data['form2'] = form
                    data['reservacion'] = reservacion = ReservacionPersonaPoli.objects.get(pk=request.GET['id'])
                    template = get_template("adm_revisareserva/modal/finalizar.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            elif action == 'ficha':
                try:
                    data['title'] = 'Reserva | Ficha'
                    data['reservacion'] = reservacion = ReservacionPersonaPoli.objects.get(pk=request.GET['id'])
                    data['detalles'] = reservacion.actividad.implementos()
                    data['observaciones'] = reservacion.observacionreservacionpersona_set.filter(status=True).order_by('fecha_creacion')
                    return render(request, "adm_revisareserva/ficha.html", data)
                except Exception as ex:
                    pass

            elif action == 'admreservar':
                try:
                    data['title']=u'Realizar Reserva'
                    data['perfiles']=TIPOS_PERFILES
                    data['areas']=AreaPolideportivo.objects.filter(status=True)
                    data['politicas'] = politicas = PoliticaPolideportivo.objects.filter(status=True, general=True,
                                                                                         mostrar=True).order_by('id')

                    return render(request, 'adm_revisareserva/admreserva.html', data)
                except Exception as ex:
                    pass

            if action == 'cargarcalendario':
                try:
                    fecha = datetime.now().date()
                    panio = fecha.year
                    pmes = fecha.month
                    if 'mover' in request.GET:
                        mover = request.GET['mover']

                        if mover == 'anterior':
                            mes = int(request.GET['mes'])
                            anio = int(request.GET['anio'])
                            pmes = mes - 1
                            if pmes == 0:
                                pmes = 12
                                panio = anio - 1
                            else:
                                panio = anio

                        elif mover == 'proximo':
                            mes = int(request.GET['mes'])
                            anio = int(request.GET['anio'])
                            pmes = mes + 1
                            if pmes == 13:
                                pmes = 1
                                panio = anio + 1
                            else:
                                panio = anio

                    id = request.GET['idactividad']
                    s_anio = panio
                    s_mes = pmes
                    s_dia = 1
                    data['mes'] = MESES_CHOICES[s_mes - 1]
                    data['ws'] = [0, 7, 14, 21, 28, 35]
                    lista = {}
                    listahorarios = []
                    data['actividad'] = actividad = ActividadPolideportivo.objects.get(pk=id)
                    data['horarios'] = horarios = HorarioActividadPolideportivo.objects.filter(status=True,
                                                                                               actividad_id=id,
                                                                                               # mostrar=True,
                                                                                               fechafin__gte=hoy,
                                                                                               )

                    # diasreserva = actividad.area.numdias
                    # if hoy.month != s_mes:
                    #     mes = hoy.month
                    #     mesubicado = s_mes
                    #     while mesubicado > mes:
                    #         numsinhorario = 0
                    #         msiguiente = calendar.monthrange(s_anio, mes)
                    #         rango = range(1, int(msiguiente[1] + 1), 1)
                    #         if mes == hoy.month:
                    #             rango = range(int(hoy.day), int(msiguiente[1] + 1), 1)
                    #         for dia in rango:
                    #             ban = False
                    #             fecha = date(s_anio, mes, dia)
                    #             numdia = fecha.weekday() + 1
                    #             for horario in horarios:
                    #                 if horario.dia == numdia:
                    #                     ban = True
                    #             if ban == False:
                    #                 numsinhorario += 1
                    #         diasreserva += numsinhorario
                    #         if mes == hoy.month:
                    #             diasreserva -= msiguiente[1] - (hoy.day - 1)
                    #         else:
                    #             diasreserva = diasreserva - msiguiente[1]
                    #
                    #         mes += 1

                    for i in range(1, 43, 1):
                        dia = {i: 'no'}
                        lista.update(dia)
                    comienzo = False
                    fin = False
                    for i in lista.items():
                        try:
                            fecha = date(s_anio, s_mes, s_dia)
                            if fecha.isoweekday() == i[0] and fin is False and comienzo is False:
                                comienzo = True
                        except Exception as ex:
                            pass
                        if comienzo:
                            try:
                                fecha = date(s_anio, s_mes, s_dia)
                            except Exception as ex:
                                fin = True
                        if comienzo and fin is False:
                            dia = {i[0]: s_dia}
                            lista.update(dia)
                            if horarios:
                                data['ultimafecha'] = ultimafecha = horarios.order_by('fechafin').last()
                                data['year'] = ultimafecha.fechafin.year

                            listhorario = []
                            sinhorario = True
                            puedereservar = True #Edite para dar acceso a reservar los dias anteriores. POR DEAULT ES FALSE
                            numerodia = fecha.weekday() + 1
                            # if fecha >= hoy: ENVIABA DICCIONARIO DESDE LA FECHA ACTUAL
                            turnos = 0
                            listturnos = []
                            for horario in horarios:
                                if horario.dia == numerodia:
                                    listhorario.append(horario.id)
                            if listhorario:
                                sinhorario = False
                                filter = horarios.filter(status=True, pk__in=listhorario)
                                for horario in filter:
                                    listturnos.append(horario.turno.comienza)
                                    turnos += horario.cupos_reservados(fecha)

                            monthRange = calendar.monthrange(s_anio, s_mes)
                            #VALIDA FECHAS A RESERVAR CUMPLIENDO CON LA PLANIFICACION CREADA
                            # if s_dia <= monthRange[1] and not sinhorario:
                            if not sinhorario:
                                ordenadas = []
                                ordenadas = sorted(listturnos, reverse=True)
                                if ordenadas[0] > datetime.now().time() or fecha > datetime.now().date():
                                    puedereservar = True
                            diccionario = {'dia': s_dia, 'turnos': turnos, 'listahorario': listhorario,
                                           'sinhorario': sinhorario, 'fecha': fecha, 'puedereservar': puedereservar}
                            listahorarios.append(diccionario)
                            s_dia += 1
                    data['dias_mes'] = lista
                    data['dwnm'] = ['lunes', 'martes', 'miercoles', 'jueves', 'viernes', 'sabado', 'domingo']
                    data['dwn'] = [1, 2, 3, 4, 5, 6, 7]
                    data['daymonth'] = 1
                    data['s_anio'] = s_anio
                    data['s_mes'] = s_mes
                    data['lista'] = lista
                    data['hoy_dia'] = hoy.day
                    data['hoy_mes'] = hoy.month
                    data['idperfil'] = int(request.GET['idperfil'])
                    data['listahorarios'] = listahorarios
                    data['fechainicio'] = date(datetime.now().year, 1, 1)
                    data['fechafin'] = datetime.now().date()
                    template = get_template("adm_revisareserva/calendario.html")
                    return JsonResponse({"result": 'ok', 'data': template.render(data)})
                except Exception as ex:
                    pass

            if action == 'selturnos':
                try:
                    data['perfil']=perfil=PerfilUsuario.objects.get(id=int(request.GET['idperfil']))
                    actividad = ActividadPolideportivo.objects.get(pk=request.GET['actividadid'])
                    if ReservacionPersonaPoli.objects.filter(status=True, actividad=actividad, estado__in=[1, 2],persona=persona).exists():
                        return JsonResponse({"result": False, "reservado": True,"mensaje": u"Usuario ya cuenta con una reservación de esta actividad en proceso, por favor finalice o cancele la pendiente para iniciar otra."})

                    horariosid = request.GET.getlist("listaid[]")
                    if not type(eval(horariosid[0])) is int:
                        horariosid = list(eval(horariosid[0]))
                    data['fecha'] = request.GET['fecha']
                    data['actividad'] = actividad
                    data['horarios'] = horario = HorarioActividadPolideportivo.objects.filter(id__in=horariosid,
                                                                                              status=True).order_by(
                        'id')
                    data['horariodia'] = horario.first()
                    data['familiares'] = PersonaDatosFamiliares.objects.filter(status=True, persona=persona)
                    data['tipoterceros'] = TIPO_TERCERO
                    data['secciones'] = SeccionPolideportivo.objects.filter(status=True, area=actividad.area,
                                                                            mostrar=True).order_by('id')
                    template = get_template("adm_revisareserva/modal/formreserva.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            if action == 'inforeserva':
                try:
                    id=int(request.GET['id'])
                    data['reserva']=reserva=ReservacionTurnosPoli.objects.get(pk=id)
                    data['reservafamiliares']=familiares=ReservacionTercerosPoli.objects.filter(status=True,reservacion=reserva)
                    data['cantidadreserva']=reserva.fechareservacion.reservacion.cantidad+len(familiares)+1
                    data['title']=reserva.turno.fechainicio
                    template=get_template('alu_reservapolideportivo/modal/inforeserva.html')
                    return JsonResponse({'result':'ok','data':template.render(data)})
                except Exception as ex:
                    pass

            if action == 'listactividades':
                try:
                    lista = []
                    idarea=int(request.GET['idarea'])
                    if not es_instructor:
                        actividadespoli=ActividadPolideportivo.objects.filter(status=True, area_id=idarea)
                        for actividadpoli in actividadespoli:
                            lista.append([actividadpoli.id, actividadpoli.nombre])
                    else:
                        actividadespoli=InstructorActividadPolideportivo.objects.filter(status=True, instructor__persona=persona, actividad__area_id=idarea).order_by('actividad').distinct('actividad')
                        for actividadpoli in actividadespoli:
                            lista.append([actividadpoli.actividad.id, actividadpoli.actividad.nombre])
                    data = {"result": "ok", "lista": lista}
                    return JsonResponse(data)
                except Exception as ex:
                    pass
        else:
            try:
                data['title'] = u'Reservas en Polideportivo'
                estreserva, perfil, instructor, area, actividad, desde, hasta, search, filtro, url_vars = request.GET.get('estreserva',''), \
                                                                                                         request.GET.get('perfil',''), \
                                                                                                         request.GET.get('instructor', ''), \
                                                                                                          request.GET.get('area', ''), \
                                                                                                          request.GET.get('actividad', ''), \
                                                                                                          request.GET.get('desde', ''), \
                                                                                                          request.GET.get('hasta', ''),\
                                                                                                          request.GET.get('s', ''), Q(status=True, planificacion__isnull=True), ''
                areas = AreaPolideportivo.objects.filter(status=True)
                if es_instructor and not request.user.has_perm('sga.puede_revisar_total_reservas') or instructor:
                    idpersona=persona.id
                    if instructor:
                        data['instructor'] = idinstructor = int(encrypt(instructor))
                        url_vars += "&instructor={}".format(instructor)
                        idpersona=idinstructor
                    instructoresactividad = InstructorActividadPolideportivo.objects.filter(status=True, instructor__persona_id=idpersona)
                    actividades=[]
                    for instructoractividad in instructoresactividad:
                        actividades.append(instructoractividad.actividad)
                    filtro = filtro & Q(actividad__in=actividades)
                    if not instructor:
                        areas=instructoresactividad.order_by('actividad__area').distinct('actividad__area')

                if perfil:
                    data['perfil'] = perfil = int(perfil)
                    url_vars += "&perfil={}".format(perfil)
                    if perfil == 1:
                        filtro= filtro & Q(perfil__inscripcion_id__gte=0)
                    elif perfil == 2:
                        filtro = filtro & Q(perfil__administrativo_id__gte=0)
                    elif perfil == 3:
                        filtro = filtro & Q(perfil__profesor_id__gte=0)
                    elif perfil == 4:
                        filtro = filtro & Q(perfil__externo_id__gte=0)

                if estreserva:
                    data['estreserva'] = estreserva = int(estreserva)
                    url_vars += "&estreserva={}".format(estreserva)
                    filtro = filtro & Q(estado=estreserva)

                if area:
                    data['area'] = idarea = int(area)
                    url_vars += f"&area={idarea}"
                    filtro = filtro & Q(area_id=idarea)

                if actividad:
                    data['actividad'] = idactividad = int(actividad)
                    url_vars += f"&actividad={idactividad}"
                    filtro = filtro & Q(actividad_id=idactividad)

                if desde:
                    data['desde'] = desde
                    url_vars += "&desde={}".format(desde)
                    filtro = filtro & Q(finicialreserva__gte=desde)

                if hasta:
                    data['hasta'] = hasta
                    url_vars += "&hasta={}".format(hasta)
                    filtro = filtro & Q(finicialreserva__lte=hasta)

                if search:
                    s = search.split(' ')
                    if len(s) == 1:
                        filtro = filtro & (Q(persona__nombres__icontains=search) | Q(persona__cedula__icontains=search))
                    if len(s) >= 2:
                        filtro = filtro & (Q(persona__apellido1__icontains=s[0]) & Q(persona__apellido2__icontains=s[1]))
                    url_vars += '&s=' + search
                    data['search'] = search

                listado = ReservacionPersonaPoli.objects.filter(filtro).order_by('-id')

                # if perfil:
                #     ids = []
                #     data['perfil'] = perfil = int(perfil)
                #     url_vars += "&perfil={}".format(perfil)
                #     for list in listado:
                #         if perfil == 1 and list.perfil.es_estudiante():
                #             ids.append(list.id)
                #         elif perfil == 2 and list.perfil.es_administrativo():
                #             ids.append(list.id)
                #         elif perfil == 3 and list.perfil.es_profesor():
                #             ids.append(list.id)
                #         elif perfil == 4 and list.perfil.es_externo():
                #             ids.append(list.id)
                #         elif perfil == 5:
                #             ids.append(list.id)
                #     listado=listado.filter(id__in=ids)

                paging = MiPaginador(listado, 20)
                p = 1
                try:
                    paginasesion = 1
                    if 'paginador' in request.session:
                        paginasesion = int(request.session['paginador'])
                    if 'page' in request.GET:
                        p = int(request.GET['page'])
                    else:
                        p = paginasesion
                    try:
                        page = paging.page(p)
                    except:
                        p = 1
                    page = paging.page(p)
                except:
                    page = paging.page(p)
                request.session['paginador'] = p
                data['paging'] = paging
                data['usuario']=usuario
                data['rangospaging'] = paging.rangos_paginado(p)
                data['page'] = page
                data["url_vars"] = url_vars
                data['listado'] = page.object_list
                data['email_domain'] = EMAIL_DOMAIN
                data['estados'] = ESTADO_RESERVA_POLIDEPORTIVO
                data['areas']=areas
                data['perfiles'] = PerfilesActividad.objects.filter(status=True).order_by('perfil').distinct('perfil')
                data['instructores']=InstructorActividadPolideportivo.objects.filter(status=True, activo=True).order_by('instructor__persona_id').distinct('instructor__persona_id')
                #CONTADOR
                data['contpendientes'] = len(listado.filter(estado=1))
                data['contreservados'] = len(listado.filter(estado=2))
                data['contanulados'] = len(listado.filter(estado=3))
                data['contfinalizados'] = len(listado.filter(estado=4))
                data['total']=len(listado)
                if 'exportar_excel' in request.GET:
                    wb = openxl.Workbook()
                    wb["Sheet"].title = "Reporte_Reservas"
                    ws = wb.active
                    style_title = openxlFont(name='Arial', size=16, bold=True)
                    style_cab = openxlFont(name='Arial', size=10, bold=True)
                    alinear = alin(horizontal="center", vertical="center")
                    response = HttpResponse(content_type="application/ms-excel")
                    response[
                        'Content-Disposition'] = 'attachment; filename=Reporte de reservaciones polideportivo' + '-' + random.randint(
                        1, 10000).__str__() + '.xlsx'
                    ws.column_dimensions['B'].width = 25
                    ws.column_dimensions['C'].width = 15
                    ws.column_dimensions['D'].width = 20
                    ws.column_dimensions['E'].width = 10
                    ws.column_dimensions['F'].width = 25
                    ws.column_dimensions['G'].width = 20
                    ws.merge_cells('A1:H1')
                    ws['A1'] = 'REPORTE DE RESERVAS DE POLIDEPORTIVO'
                    celda1 = ws['A1']
                    celda1.font = style_title
                    celda1.alignment = alinear

                    columns = [u"N°", u"NOMBRES Y APELLIDOS",
                               u"CÉDULA", u"PERFIL",
                               u"¿ASISTIO?", u"CARRERA",
                               u"MODALIDAD", u"ESTADO",
                               "FECHA DE RESERVA", "HORA DE RESERVA",
                               "AREA", "ACTIVIDAD", "DISCIPLINA", "INSTRUCTORES",
                               ]
                    row_num = 3
                    for col_num in range(0, len(columns)):
                        celda = ws.cell(row=row_num, column=(col_num + 1), value=columns[col_num])
                        celda.font = style_cab

                    mensaje = 'NO REGISTRA'
                    row_num = 4
                    numero = 0
                    for reserva in listado:
                        numero += 1
                        ws.cell(row=row_num, column=1, value=numero)
                        ws.cell(row=row_num, column=2, value=str(reserva.persona))
                        ws.cell(row=row_num, column=3, value=str(reserva.persona.cedula))
                        ws.cell(row=row_num, column=4, value=str(reserva.perfil))
                        ws.cell(row=row_num, column=5, value='SI' if reserva.asistio else 'NO')
                        ws.cell(row=row_num, column=6, value=str(reserva.perfil.inscripcion.carrera) if reserva.perfil.es_estudiante() else mensaje)
                        ws.cell(row=row_num, column=7, value=str(reserva.perfil.inscripcion.modalidad) if reserva.perfil.es_estudiante() else mensaje)
                        ws.cell(row=row_num, column=8, value=reserva.get_estado_display())
                        ws.cell(row=row_num, column=9, value=str(reserva.finicialreserva))
                        ws.cell(row=row_num, column=10, value=str(reserva.turno_reservado().turno.turno.comienza)+' a '+str(reserva.turno_reservado().turno.turno.termina))
                        ws.cell(row=row_num, column=11, value=str(reserva.area))
                        ws.cell(row=row_num, column=12, value=str(reserva.actividad))
                        ws.cell(row=row_num, column=13, value=str(reserva.actividad.disciplina))
                        columna = 13
                        for ins in reserva.actividad.instructores():
                            columna += 1
                            ws.cell(row=row_num, column=columna, value=ins.instructor.persona.nombre_completo_minus())
                        row_num += 1
                    wb.save(response)
                    return response
                return render(request, 'adm_revisareserva/view.html', data)
            except Exception as ex:
                print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
                pass