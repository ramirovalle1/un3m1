#!/usr/bin/env python
# -*- coding: utf-8 -*-
import os
import sys
import time

import openpyxl

# SITE_ROOT = os.path.dirname(os.path.realpath(__file__))


YOUR_PATH = os.path.dirname(os.path.realpath(__file__))
# print(f"YOUR_PATH: {YOUR_PATH}")
SITE_ROOT = os.path.dirname(os.path.dirname(YOUR_PATH))
SITE_ROOT = os.path.join(SITE_ROOT, '')
print(f"SITE_ROOT: {SITE_ROOT}")
your_djangoproject_home = os.path.split(SITE_ROOT)[0]
# print(f"your_djangoproject_home: {your_djangoproject_home}")
sys.path.append(your_djangoproject_home)

from django.core.wsgi import get_wsgi_application
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "settings")
application = get_wsgi_application()
from inno.models import *
from sga.models import *
from sagest.models import *
from posgrado.models import GrupoExamenMsc
from sga.funciones import convertir_fecha,convertir_hora
from postulate.models import PersonaAplicarPartida,CalificacionPostulacion
import xlrd
from moodle import moodle
from bd.models import InventarioOpcionSistema
from sga.My_Model.SubirMatrizSENESCYT import My_HistorialProcesoSubirMatrizInscripcion, \
    My_HistorialSubirMatrizInscripcion, My_SubirMatrizInscripcion
from oma.models import *
from urllib.parse import urlencode
from urllib.request import urlopen, Request
import json
import re
import xlwt
from xlwt import *
from django.http import HttpResponse
import warnings
warnings.filterwarnings('ignore', message='Unverified HTTPS request')
# try:
#     for horario in HorarioTutoriaAcademica.objects.filter(status=True):
#         if ProfesorMateria.objects.filter(status=True, profesor=horario.profesor,
#                                           materia__nivel__periodo=horario.periodo,
#                                           activo=True).exists():
#             cant1 = ProfesorMateria.objects.values('id').filter(status=True, profesor=horario.profesor,
#                                                                 materia__nivel__periodo=horario.periodo,
#                                                                 activo=True,
#                                                                 tipoprofesor__id__in=[8]).count()
#             cant2 = ProfesorMateria.objects.values('id').filter(status=True, profesor=horario.profesor,
#                                                                 materia__nivel__periodo=horario.periodo,
#                                                                 activo=True).count()
#             if cant1 == cant2:
#                 print('Registro eliminado %s' % horario)
#                 horario.status=False
#                 horario.save()
#
#
# except Exception as ex:
#     print('error: %s' % ex)

# try:
#
#     miarchivo = openpyxl.load_workbook("insertar_actividades.xlsx")
#     lista = miarchivo.get_sheet_by_name('Hoja1')
#     totallista = lista.rows
#     a=0
#     periodo=Periodo.objects.get(id=119)
#     for filas in totallista:
#         a += 1
#         if a > 1:
#             cedula = str(filas[0].value).strip() if filas[0].value else None
#             if Persona.objects.filter(Q(cedula=cedula) | Q(pasaporte=cedula)| Q(ruc=cedula) ,status=True).exists():
#                 persona=Persona.objects.filter( Q(cedula=cedula) | Q(pasaporte=cedula)| Q(ruc=cedula) ,status=True)[0]
#                 if persona.es_profesor():
#                     profesor=Profesor.objects.filter(status=True,persona=persona)[0]
#                     print(u"%s" % profesor)
#                     if profesor.distributivohoraseval(periodo):
#                         distributivo= profesor.distributivohoraseval(periodo)
#                         if not DetalleDistributivo.objects.values('id').filter(status=True,distributivo=distributivo,criteriodocenciaperiodo_id=569 ).exists():
#                             insdis_58=DetalleDistributivo(distributivo=distributivo,criteriodocenciaperiodo_id=569,horas=8)
#                             insdis_58.save()
#                         else:
#                             insdis_58=DetalleDistributivo.objects.filter(status=True, distributivo=distributivo,criteriodocenciaperiodo_id=569)[0]
#                             insdis_58.horas=8
#                             insdis_58.save()
#
#                         if not DetalleDistributivo.objects.values('id').filter(status=True,distributivo=distributivo,criteriodocenciaperiodo_id=572 ).exists():
#                             insdis_105=DetalleDistributivo(distributivo=distributivo,criteriodocenciaperiodo_id=572,horas=12)
#                             insdis_105.save()
#                         else:
#                             insdis_105=DetalleDistributivo.objects.filter(status=True, distributivo=distributivo, criteriodocenciaperiodo_id=572)[0]
#                             insdis_105.horas=12
#                             insdis_105.save()
#
#                         for detalledistributivo in DetalleDistributivo.objects.filter(status=True,criteriodocenciaperiodo__isnull=False,distributivo=distributivo).order_by('-criteriodocenciaperiodo__criterio_id'):
#                             actividad1=None
#                             actividad2=None
#                             actividad3=None
#                             actividad4=None
#                             actividad5=None
#                             actividad6=None
#                             if detalledistributivo.criteriodocenciaperiodo.criterio_id==105:
#                                 if ActividadDetalleDistributivo.objects.filter(status=True,criterio=detalledistributivo, criterio__criteriodocenciaperiodo__criterio__id=105).exists():
#                                     subact=ActividadDetalleDistributivo.objects.filter(status=True,
#                                                                                 criterio=detalledistributivo,
#                                                                                 criterio__criteriodocenciaperiodo__criterio__id=105).delete()
#                                 if not ActividadDetalleDistributivo.objects.filter(status=True,criterio=detalledistributivo,
#                                                                                    criterio__criteriodocenciaperiodo__criterio__id=105,
#                                                                                    nombre=u"Corregir, calificar y comentar (retroalimentación) las actividades programadas en el aula virtual para los estudiantes.",
#                                                                                    ).exists():
#                                     actividad1 = ActividadDetalleDistributivo(criterio=detalledistributivo,
#                                                                              nombre=u"Corregir, calificar y comentar (retroalimentación) las actividades programadas en el aula virtual para los estudiantes.",
#                                                                              desde=convertir_fecha('29-11-2021'),
#                                                                              hasta=convertir_fecha('24-03-2022'),
#                                                                              horas=10,
#                                                                              vigente=True)
#                                     actividad1.save()
#                                     print(u"Inserta atividad 1 %s"%actividad1)
#                                 if not ActividadDetalleDistributivo.objects.filter(status=True,
#                                                                                        criterio=detalledistributivo,
#                                                                                        criterio__criteriodocenciaperiodo__criterio__id=105,
#                                                                                        nombre=u"Participar en la planificación, ejecución y cierre del proceso de evaluación académica del estudiante conforme le sea designado (Revisión de los recursos y actividades cargadas en el aula virtual).",
#                                                                                        ).exists():
#                                     actividad2 = ActividadDetalleDistributivo(criterio=detalledistributivo,
#                                                                              nombre=u"Participar en la planificación, ejecución y cierre del proceso de evaluación académica del estudiante conforme le sea designado (Revisión de los recursos y actividades cargadas en el aula virtual).",
#                                                                              desde=convertir_fecha('29-11-2021'),
#                                                                              hasta=convertir_fecha('24-03-2022'),
#                                                                              horas=1,
#                                                                              vigente=True)
#                                     actividad2.save()
#                                     print(u"Inserta atividad 2 %s" % actividad2)
#                                 if not ActividadDetalleDistributivo.objects.filter(status=True,
#                                                                                        criterio=detalledistributivo,
#                                                                                        criterio__criteriodocenciaperiodo__criterio__id=105,
#                                                                                        nombre=u"Asistir a todas las reuniones presenciales o virtuales a las que sea convocado. (Coordinación de trabajo y seguimiento con autor y/o director de carrera).",
#                                                                                        ).exists():
#                                     actividad3 = ActividadDetalleDistributivo(criterio=detalledistributivo,
#                                                                              nombre=u"Asistir a todas las reuniones presenciales o virtuales a las que sea convocado. (Coordinación de trabajo y seguimiento con autor y/o director de carrera).",
#                                                                              desde=convertir_fecha('29-11-2021'),
#                                                                              hasta=convertir_fecha('24-03-2022'),
#                                                                              horas=1,
#                                                                              vigente=True)
#                                     actividad3.save()
#                                     print(u"Inserta atividad 3 %s" % actividad2)
#                             if detalledistributivo.criteriodocenciaperiodo.criterio_id == 58:
#                                 if ActividadDetalleDistributivo.objects.filter(status=True,criterio=detalledistributivo,
#                                                                                    criterio__criteriodocenciaperiodo__criterio__id=58
#                                                                                    ).exists():
#                                     subact=ActividadDetalleDistributivo.objects.filter(status=True,
#                                                                                 criterio=detalledistributivo,
#                                                                                 criterio__criteriodocenciaperiodo__criterio__id=58).delete()
#
#                                 if not ActividadDetalleDistributivo.objects.filter(status=True,criterio=detalledistributivo,
#                                                                                    criterio__criteriodocenciaperiodo__criterio__id=58,
#                                                                                    nombre=u"Realizar el respectivo seguimiento a los estudiantes mediante la plataforma virtual siguiendo los lineamientos que le hayan sido otorgados."
#                                                                                    ).exists():
#                                     actividad4 = ActividadDetalleDistributivo(criterio=detalledistributivo,
#                                                                              nombre=u"Realizar el respectivo seguimiento a los estudiantes mediante la plataforma virtual siguiendo los lineamientos que le hayan sido otorgados.",
#                                                                              desde=convertir_fecha('29-11-2021'),
#                                                                              hasta=convertir_fecha('24-03-2022'),
#                                                                              horas=5,
#                                                                              vigente=True)
#                                     actividad4.save()
#                                     print(u"Inserta atividad 4 %s" % actividad4)
#                                 if not ActividadDetalleDistributivo.objects.filter(status=True,
#                                                                                        criterio=detalledistributivo,
#                                                                                        criterio__criteriodocenciaperiodo__criterio__id=58,
#                                                                                    nombre=u"Completar y entregar los formatos, matrices y documentos que le sean requeridos como parte del seguimiento académico o gestiones asignadas.",
#                                                                                        ).exists():
#                                     actividad5 = ActividadDetalleDistributivo(criterio=detalledistributivo,
#                                                                              nombre=u"Completar y entregar los formatos, matrices y documentos que le sean requeridos como parte del seguimiento académico o gestiones asignadas.",
#                                                                              desde=convertir_fecha('29-11-2021'),
#                                                                              hasta=convertir_fecha('24-03-2022'),
#                                                                              horas=2,
#                                                                              vigente=True)
#                                     actividad5.save()
#                                     print(u"Inserta atividad 5 %s" % actividad5)
#                                 if not ActividadDetalleDistributivo.objects.filter(status=True,
#                                                                                        criterio=detalledistributivo,
#                                                                                        criterio__criteriodocenciaperiodo__criterio__id=58,
#                                                                                        nombre=u"Atender las consultas presenciales y virtuales que le formulen los estudiantes en cuanto a los temas que le correspondan.",
#                                                                                        ).exists():
#                                     actividad6 = ActividadDetalleDistributivo(criterio=detalledistributivo,
#                                                                              nombre=u"Atender las consultas presenciales y virtuales que le formulen los estudiantes en cuanto a los temas que le correspondan.",
#                                                                              desde=convertir_fecha('29-11-2021'),
#                                                                              hasta=convertir_fecha('24-03-2022'),
#                                                                              horas=1,
#                                                                              vigente=True)
#                                     actividad6.save()
#                                     print(u"Inserta atividad 6 %s" % actividad6)
#                                 if actividad6:
#                                     actividad6.actualiza_padre()
# except Exception as ex:
#         print('error: %s' % ex)


# try:
#
#     miarchivo = openpyxl.load_workbook("insertar_actividades.xlsx")
#     lista = miarchivo.get_sheet_by_name('Hoja2')
#     totallista = lista.rows
#     a=0
#     periodo=Periodo.objects.get(id=113)
#     for filas in totallista:
#         a += 1
#         if a > 1:
#             cedula = str(filas[0].value).strip() if filas[0].value else None
#             if Persona.objects.filter(Q(cedula=cedula) | Q(pasaporte=cedula)| Q(ruc=cedula) ,status=True).exists():
#                 persona=Persona.objects.filter( Q(cedula=cedula) | Q(pasaporte=cedula)| Q(ruc=cedula) ,status=True)[0]
#                 if persona.es_profesor():
#                     profesor=Profesor.objects.filter(status=True,persona=persona)[0]
#                     print(u"%s" % profesor)
#                     if profesor.distributivohoraseval(periodo):
#                         distributivo= profesor.distributivohoraseval(periodo)
#                         distributivo.categoria_id=9
#                         distributivo.nivelcategoria_id=2
#                         distributivo.save()
#                         print('Registro cambiado: %s' % distributivo)
# except Exception as ex:
#         print('error: %s' % ex)


#
# try:
#     periodo=Periodo.objects.get(id=119)
#     idp =HorarioTutoriaAcademica.objects.values_list('profesor_id').filter(status=True, periodo=periodo).distinct()
#     for dis in ProfesorDistributivoHoras.objects.filter(status=True, periodo=periodo, activo=True,
#                                                         detalledistributivo__criteriodocenciaperiodo__criterio_id__in=[7],
#                                                         profesor_id__in=idp).distinct():
#         print('DOCENTE .... ---  .....: %s' % dis.profesor)
#         if DetalleDistributivo.objects.filter(distributivo__profesor=dis.profesor,
#                                                distributivo__periodo=periodo,criteriodocenciaperiodo__criterio_id__in=[7]).exists():
#             totalhoras=int(DetalleDistributivo.objects.filter(distributivo__profesor=dis.profesor,
#                                                    distributivo__periodo=periodo,criteriodocenciaperiodo__criterio_id__in=[7]).aggregate(total=Sum('horas'))['total'])
#
#             if HorarioTutoriaAcademica.objects.filter(status=True, profesor=dis.profesor,periodo=periodo).exists():
#                 horarios = HorarioTutoriaAcademica.objects.filter(status=True, profesor=dis.profesor,periodo=periodo)
#                 totalhoras=horarios.count()
#                 fecha = convertir_fecha('01-03-2022')
#                 for x in range(31):
#                     if fecha != convertir_fecha('01-04-2022'):
#                         for horario in horarios:
#                             if horario.dia == fecha.isoweekday():
#                                 cantasissemana=RegistroClaseTutoriaDocente.objects.filter(numerosemana=fecha.isocalendar()[1], horario__profesor=dis.profesor).aggregate(total=Count('id'))['total']
#                                 if totalhoras > cantasissemana:
#                                     if not RegistroClaseTutoriaDocente.objects.filter(horario=horario,
#                                                                                    numerosemana=fecha.isocalendar()[1],
#                                                                                    fecha__date=fecha).exists():
#                                         clasetutoria = RegistroClaseTutoriaDocente(horario=horario,
#                                                                                    numerosemana=fecha.isocalendar()[1],
#                                                                                    fecha=fecha)
#                                         clasetutoria.save()
#                                         print('Se ingresa tutoria: %s' % clasetutoria)
#                     fecha=fecha + timedelta(days=1)
#             else:
#                 print('No tiene horario: %s' % dis)
# except Exception as ex:
#         print('error: %s' % ex)



# from django.db import connections
# cursor = connections['moodle_db'].cursor()
# tareas = TareaSilaboSemanal.objects.values_list('idtareamoodle', flat=True).filter(id__in=[15376,5185])
#
# foros = ForoSilaboSemanal.objects.values_list('idforomoodle', flat=True).filter(id__in=[1827, 606])
#
# todas=list(tareas)+list(foros)
#
#
# query = """SELECT gra.finalgrade, gra.itemid
#                 FROM mooc_grade_grades gra
#                 inner join mooc_grade_items item ON gra.itemid=item.id
#                 WHERE gra.finalgrade IS NOT NULL
#                 AND gra.userid=%s and item.courseid=%s
#                 AND item.iteminstance = (
#                 SELECT cm.instance
#                 FROM mooc_course_modules cm
#                 INNER JOIN mooc_course co ON co.id=cm.course
#                 WHERE co.id=%s AND cm.id in (%s))
#                             """ % (materiaasignada.matricula.inscripcion.persona.idusermoodle,
#                                    materiaasignada.materia.idcursomoodle,
#                                    materiaasignada.materia.idcursomoodle,
#                                    tuple(todas))
# cursor.execute(query)
# results = cursor.fetchall()
#
# print(todas)


# periodo=Periodo.objects.get(id=119)
# from django.db import connections
# cursor = connections['moodle_db'].cursor()
# for materia in Materia.objects.filter(status=True, nivel__periodo=periodo,
#                                       asignaturamalla__malla__carrera__coordinacion__in=[1,2,3,4,5],
#                                       horarioexamen__status=True):
#     for materiaasignada in materia.asignados_a_esta_materia():
#         sql = """ SELECT TO_TIMESTAMP(att.timestart), TO_TIMESTAMP(att.timefinish)
#                 FROM mooc_quiz_attempts att
#                 INNER JOIN mooc_quiz quiz ON quiz.id=att.quiz
#                 INNER JOIN mooc_course cou ON cou.id=quiz.course
#                 INNER JOIN  mooc_grade_items gi ON gi.courseid=cou.id AND gi.itemmodule='quiz' AND quiz.id=gi.iteminstance
#                 INNER JOIN  mooc_grade_categories ct ON ct.courseid=cou.id
#                 WHERE cou.id=%s AND att.userid=%s AND ct.fullname='EX1'
#                 AND ct.courseid=%s AND ct.depth=2 AND ct.id=gi.categoryid
#                                     """ % ( materia.idcursomoodle, materiaasignada.matricula.inscripcion.persona.idusermoodle,materia.idcursomoodle)
#         cursor.execute(sql)
#         row = cursor.fetchall()
#         lista = []
#         for r in row:
#             if r:
#                 fechainicio = r[0]
#                 horario = HorarioExamen.objects.filter(status=True, materia=materia, detallemodelo__nombre="EX1")[0]
#                 if horario.fecha == fechainicio.date():
#                     if fechainicio.date() < convertir_fecha('17-01-2022'):
#                         if not SesionZoom.objects.filter(materiaasignada=materiaasignada, status=True, fecha=horario.fecha,  modulo=2).exists():
#                             zoom = SesionZoom(
#                                 materiaasignada=materiaasignada,
#                                 modulo=2,
#                                 fecha=horario.fecha,
#                                 hora=fechainicio.time()
#                             )
#                             zoom.save()
#                             print("Inserta horario EX1 %s %s %s" % ( materiaasignada.matricula.inscripcion.persona, horario.fecha, fechainicio.time()))
                        # else:
                        #     if SesionZoom.objects.filter(materiaasignada=materiaasignada, status=True, fecha=horario.fecha, modulo=2,usuario_creacion_id=1,fecha_creacion__date=convertir_fecha("29-07-2021")).exists():
                        #         zoom = SesionZoom.objects.filter(materiaasignada=materiaasignada, status=True, fecha=horario.fecha, modulo=2,usuario_creacion_id=1,fecha_creacion__date=convertir_fecha("29-07-2021"))[0]
                        #         zoom.hora = fechainicio.time()
                        #         zoom.horaultima = fechainicio.time()
                        #         zoom.save()
                        #         print("MODIFICA horario EX1 %s %s %s" % (
                        #         materiaasignada.matricula.inscripcion.persona, horario.fecha, fechainicio.time()))


# periodo=Periodo.objects.get(status=True,id=113)
# for carrera in Carrera.objects.filter(status=True, coordinacion__id__in=[1,2,3,4,5], modalidad__in=[1,2]):
#     for tipo in TipoRecurso.objects.filter(status=True):
#         if not ConfiguracionRecurso.objects.filter(status=True, periodo=periodo,
#                                                    tiporecurso=tipo,carrera=carrera).exists():
#             config=ConfiguracionRecurso(status=True, periodo=periodo,
#                                                 tiporecurso=tipo, carrera=carrera)
#             config.save()
#             print("---------- %s ---------" % config)
#             if tipo.id in [2,3]:
#                 config.formato=FormatoArchivo.objects.filter(id=3)
#                 print("formato PDF")
#             elif tipo.id == 1:
#                 config.formato=FormatoArchivo.objects.filter(id__in=[2,3])
#                 print("formato POWER POINT")
#                 print("formato PDF")


#
# periodo=Periodo.objects.get(id=119)
# cont=0
# mat_id = HorarioExamenDetalle.objects.values_list('horarioexamen__materia__id',flat=True).filter(status=True,
#                                                                                                  horarioexamen__materia__nivel__periodo=periodo,
#                                                                                                  horarioexamen__fecha=convertir_fecha('15-03-2022'),
#                                                                                                  horainicio__range=(convertir_hora('07:00:00'),
#                                                                                                                     convertir_hora('07:30:00')),
#                                                                                                  horarioexamen__detallemodelo__nombre="EX2")
# for materia in Materia.objects.filter(status=True, nivel__periodo=periodo,id__in=mat_id,
#                                       asignaturamalla__malla__carrera__coordinacion__in=[1,2,3,4,5],
#                                       horarioexamen__status=True):
#     horario = HorarioExamen.objects.filter(status=True, materia=materia,fecha=convertir_fecha('15-03-2022'), detallemodelo__nombre="EX2")[0]
#     detalle = horario.horarioexamendetalle_set.filter(status=True)
#     if detalle:
#         horainiciod = detalle[0].horainicio
#         if horario.fecha == convertir_fecha('15-03-2022'):
#             if horainiciod >= convertir_hora('07:00:00') and horainiciod <= convertir_hora('07:30:00'):
#                 for materiaasignada in materia.asignados_a_esta_materia():
#                     if not SesionZoom.objects.filter(materiaasignada=materiaasignada, status=True, fecha=horario.fecha, modulo=2).exists():
#                         zoom = SesionZoom(
#                             materiaasignada=materiaasignada,
#                             modulo=2,
#                             fecha=horario.fecha,
#                             hora=horainiciod
#                         )
#                         zoom.save()
#                         cont += 1
#                         print("%s Inserta horario EX2 %s %s %s" % (cont, materiaasignada.matricula, horario.fecha, zoom.hora))



# workbook = xlrd.open_workbook("matriculas_salud.xlsx")
# sheet = workbook.sheet_by_index(0)
# linea = 1
# cedula = 0
# carrera = 0
# periodo = Periodo.objects.get(pk=122)
# sede = Sede.objects.get(pk=1)
# import time
# try:
#     for rowx in range(sheet.nrows):
#         if linea>1:
#             cols = sheet.row_values(rowx)
#             session_id = int(cols[13])
#             cedula = cols[0].strip().upper()
#             persona=None
#             idcarrera=None
#             if Persona.objects.filter(Q(cedula=cedula) | Q(pasaporte=cedula), status=True).exists():
#                 datospersona = Persona.objects.filter(Q(cedula=cedula) | Q(pasaporte=cedula), status=True)[0]
#             carrera = Carrera.objects.get(pk=idcarrera)
#             modalidad = Modalidad.objects.get(pk=int(cols[8]))
#             if Inscripcion.objects.values('id').filter(status=True,persona=persona,carrera=carrera).exists():
#                 inscripcion = Inscripcion.objects.filter(persona=persona,carrera=carrera)[0]
#
#             nivel = Nivel.objects.get(periodo=periodo, id=int(cols[12]))
#             if not inscripcion.matricula_periodo(periodo):
#                 matricula = Matricula(inscripcion=inscripcion,
#                                       nivel=nivel,
#                                       pago=False,
#                                       iece=False,
#                                       becado=False,
#                                       porcientobeca=0,
#                                       fecha=datetime.now().date(),
#                                       hora=datetime.now().time(),
#                                       fechatope=fechatope(datetime.now().date()))
#                 matricula.save()
#             else:
#                 matricula = Matricula.objects.get(inscripcion=inscripcion, nivel=nivel)
#
#             for materia in Materia.objects.filter(nivel__periodo=periodo, paralelo=cols[7].strip(), asignaturamalla__malla=mimalla.malla, asignaturamalla__malla__carrera=carrera, nivel__sesion=sesion):
#                 if not MateriaAsignada.objects.values('id').filter(matricula=matricula,materia=materia).exists():
#                     matriculas = matricula.inscripcion.historicorecordacademico_set.values('id').filter(asignatura=materia.asignatura, fecha__lt=materia.nivel.fin).count() + 1
#                     materiaasignada = MateriaAsignada(matricula=matricula,
#                                                       materia=materia,
#                                                       notafinal=0,
#                                                       asistenciafinal=0,
#                                                       cerrado=False,
#                                                       matriculas=matriculas,
#                                                       observaciones='',
#                                                       estado_id=NOTA_ESTADO_EN_CURSO)
#                     materiaasignada.save()
#                     materiaasignada.asistencias()
#                     materiaasignada.evaluacion()
#                     materiaasignada.mis_planificaciones()
#                     materiaasignada.save()
#                     print(materiaasignada)
#             matricula.actualizar_horas_creditos()
#             matricula.estado_matricula=2
#             matricula.save()
#             matricula.calcula_nivel()
#             inscripcion.actualizar_nivel()
#         print(linea, cedula, carrera)
#         linea += 1
# except Exception as ex:
#     print(ex)

# try:
#     materias = Materia.objects.filter(status=True,nivel__periodo__id=113,asignaturamalla__malla__carrera__id__in=[139, 133,22])
#     persona=Persona.objects.get(id=22977)
#     fecha=convertir_fecha_hora("10-09-2021 00:00:00")
#     cont=0
#     for materia in materias:
#         cont+=1
#         for detalle in materia.detallemodeloevaluativo().filter(nombre="EX2"):
#             if ReactivoMateria.objects.filter(materia=materia, detallemodelo=detalle):
#                 reactivomateria=ReactivoMateria.objects.filter(materia=materia, detallemodelo=detalle)[0]
#                 reactivomateria.fecha=fecha
#                 print("%s Se modifica fecha %s "%(cont,reactivomateria))
#             else:
#                 reactivomateria = ReactivoMateria(materia=materia,
#                                                               fecha=fecha,
#                                                               persona=persona,
#                                                               detallemodelo=detalle)
#                 print("%s Se crea fecha %s" % (cont,reactivomateria))
#             reactivomateria.save()
# except Exception as ex:
#     print(ex)

# try:
#     materia=Materia.objects.get(id=44277)
#     matricula=Matricula.objects.get(id=350930)
#     materia.crear_actualizar_un_estudiante_curso(moodle, 2,matricula)
# except Exception as ex:
#     print(ex)

# import xlwt
# from xlwt import *
# from django.http import HttpResponse
#
# response = HttpResponse(content_type="application/ms-excel")
# response['Content-Disposition'] = 'attachment; filename=reporte_predecesoras.xls'
# style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
# style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
# style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
# title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
# style1 = easyxf(num_format_str='D-MMM-YY')
# font_style = XFStyle()
# font_style.font.bold = True
# font_style2 = XFStyle()
# font_style2.font.bold = False
# wb = xlwt.Workbook()
# ws = wb.add_sheet('Sheetname')
# estilo = xlwt.easyxf('font: height 350, name Arial, colour_index black, bold on, italic on; align: wrap on, vert centre, horiz center;')
# ws.write_merge(0, 0, 0, 9, 'UNIVERSIDAD ESTATAL DE MILAGRO', estilo)
# output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
# nombre = "Lista" + datetime.now().strftime('%Y%m%d_%H%M%S') + ".xls"
# filename = os.path.join(output_folder, nombre)
# columns = [(u"Id", 6000),
#            (u"Materia", 6000),
#            (u"Nivel", 6000),
#            (u"Cant pre decesora", 6000),
#            ]
# row_num = 3
# for col_num in range(len(columns)):
#     ws.write(row_num, col_num, columns[col_num][0], font_style)
#     ws.col(col_num).width = columns[col_num][1]
# row_num = 4
# periodo=Periodo.objects.get(id=119)
# materias=Materia.objects.filter(status=True, nivel__periodo=periodo)
# for materia in materias:
#     if materia.asignaturamalla.nivelmalla_id>4:
#         ws.write(row_num, 0, u'%s' % materia.id, font_style2)
#         ws.write(row_num, 1, u'%s' % materia, font_style2)
#         ws.write(row_num, 2, u'%s' % materia.asignaturamalla.nivelmalla, font_style2)
#         ws.write(row_num, 3, u'%s' % materia.asignaturamalla.asignaturamallapredecesora_set.filter(status=True).count() if  materia.asignaturamalla.asignaturamallapredecesora_set.filter(status=True).exists() else 0, font_style2)
#         print(u"%s" % materia)
#         row_num += 1
# wb.save(filename)
# print("FIN: ", filename)
#




# for inscripcion in Inscripcion.objects.filter(status=True,id__in=[8247,19955,8835, 17315,17313,17310, 17297, 5725, 14497, 5373, 17336, 17283, 3684,
#                                                                 3681,
#                                                                 6122,
#                                                                 8690,
#                                                                 12417,
#                                                                 18024,
#                                                                 18021,
#                                                                 11192,
#                                                                 16133,
#                                                                 18018,
#                                                                 18019,
#                                                                 19900,
#                                                                 17322,
#                                                                 17319,
#                                                                 17306,
#                                                                 17292,
#                                                                 17291,
#                                                                 3175,
#                                                                 3531,
#                                                                 28478,
#                                                                 8092,
#                                                                 5416,
#                                                                 5437,
#                                                                 790,
#                                                                 8909,
#                                                                 9075,
#                                                                 9301,
#                                                                 16189,
#                                                                 30537,
#                                                                 17278,
#                                                                 10152,
#                                                                 29352,
#                                                                 9643,
#                                                                 20009,
#                                                                 17299,
#                                                                 17298,
#                                                                 17293,
#                                                                 5757,
#                                                                 17277,
#                                                                 17274,
#                                                                 5288,
#                                                                 3707,
#                                                                 18004,
#                                                                 17339,
#                                                                 17318,
#                                                                 17335,
#                                                                 16142,
#                                                                 18022,
#                                                                 19918,
#                                                                 20756,
#                                                                 15975,
#                                                                 19911,
#                                                                 9842,
#                                                                 17309,
#                                                                 5808,
#                                                                 19946,
#                                                                 19931,
#                                                                 19899,
#                                                                 19912,
#                                                                 13619,
#                                                                 19909,
#                                                                 20132,
#                                                                 19223,
#                                                                 5330,
#                                                                 4819,
#                                                                 5084,
#                                                                 16005,
#                                                                 18527,
#                                                                 19917,
#                                                                 19903,
#                                                                 25943,
#                                                                 16201,
#                                                                 15239,
#                                                                 19943,
#                                                                 3190,
#                                                                 17279,
#                                                                 15841,
#                                                                 9611,
#                                                                 13032,
#                                                                 17373,
#                                                                 19947,
#                                                                 17314,
#                                                                 17300,
#                                                                 9245,
#                                                                 17281,
#                                                                 17286,
#                                                                 17282,
#                                                                 17272,
#                                                                 17357,
#                                                                 17330,
#                                                                 15731,
#                                                                 3413,
#                                                                 2661,
#                                                                 5728,
#                                                                 1424,
#                                                                 19927,
#                                                                 19926,
#                                                                 16090,
#                                                                 18020,
#                                                                 19528,
#                                                                 20468,
#                                                                 18023,
#                                                                 1529,
#                                                                 17301,
#                                                                 17290,
#                                                                 17288,
#                                                                 17307,
#                                                                 80,
#                                                                 5282,
#                                                                 13370,
#                                                                 16333,
#                                                                 8979,
#                                                                 20506,
#                                                                 19690,
#                                                                 19923,
#                                                                 19898,
#                                                                 19915,
#                                                                 16110,
#                                                                 10180,
#                                                                 16036,
#                                                                 19902,
#                                                                 19993,
#                                                                 15316,
#                                                                 18722,
#                                                                 15937,
#                                                                 15920,
#                                                                 17284,
#                                                                 29351,
#                                                                 15476,
#                                                                 19920,
#                                                                 19904,
#                                                                 9896,
#                                                                 17316,
#                                                                 17312,
#                                                                 17311,
#                                                                 17295,
#                                                                 16324,
#                                                                 30532,
#                                                                 15764,
#                                                                 17275,
#                                                                 5347,
#                                                                 5346,
#                                                                 17355,
#                                                                 17328,
#                                                                 17276,
#                                                                 19919,
#                                                                 5799,
#                                                                 29353,
#                                                                 17305,
#                                                                 17287,
#                                                                 1418,
#                                                                 3832,
#                                                                 19937,
#                                                                 18998,
#                                                                 19951,
#                                                                 19938,
#                                                                 20853,
#                                                                 9903,
#                                                                 9902,
#                                                                 19914,
#                                                                 19882,
#                                                                 19941,
#                                                                 19227,
#                                                                 19907,
#                                                                 15935,
#                                                                 11040,
#                                                                 5415,
#                                                                 16309,
#                                                                 29244,
#                                                                 16149]):
#     inscripcion.malla_inscripcion()
#     print(u"%s"%inscripcion)
#
#
# import xlwt
# from xlwt import *
# from django.http import HttpResponse
#
# response = HttpResponse(content_type="application/ms-excel")
# response['Content-Disposition'] = 'attachment; filename=reporte_matriculados.xls'
# style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
# style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
# style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
# title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
# style1 = easyxf(num_format_str='D-MMM-YY')
# font_style = XFStyle()
# font_style.font.bold = True
# font_style2 = XFStyle()
# font_style2.font.bold = False
# wb = xlwt.Workbook()
# ws = wb.add_sheet('Sheetname')
# estilo = xlwt.easyxf('font: height 350, name Arial, colour_index black, bold on, italic on; align: wrap on, vert centre, horiz center;')
# ws.write_merge(0, 0, 0, 9, 'UNIVERSIDAD ESTATAL DE MILAGRO', estilo)
# output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
# nombre = "Lista" + datetime.now().strftime('%Y%m%d_%H%M%S') + ".xls"
# filename = os.path.join(output_folder, nombre)
# columns = [(u"Carrera", 6000),
#            (u"Cantidad", 6000),
#            ]
# row_num = 3
# for col_num in range(len(columns)):
#     ws.write(row_num, col_num, columns[col_num][0], font_style)
#     ws.col(col_num).width = columns[col_num][1]
# row_num = 4
# for carrera in Carrera.objects.filter(pk__in=Materia.objects.values_list('asignaturamalla__malla__carrera').filter(nivel__periodo__id__in=[112,113]).distinct()).order_by('-nombre').distinct():
#     cant=0
#     cant= Matricula.objects.values_list('inscripcion__persona__id').filter(status=True,inscripcion__status=True,estado_matricula__in=[2,3],retiradomatricula=False, nivel__periodo__id__in=[112,113],inscripcion__carrera=carrera).order_by('inscripcion__persona__id').distinct('inscripcion__persona__id').count()
#     ws.write(row_num, 0, u'%s' % carrera, font_style2)
#     ws.write(row_num, 1, u'%s' % cant, font_style2)
#     print(u"%s" % carrera)
#     row_num += 1
# wb.save(filename)
# print("FIN: ", filename)


# def agregacion_aux(matricula):
#     if matricula.inscripcion.coordinacion_id != 9 and matricula.inscripcion.coordinacion_id != 7:
#         cantidad_seleccionadas = 0
#         cursor = connections['default'].cursor()
#         sql = "select am.nivelmalla_id, count(am.nivelmalla_id) as cantidad_materias_seleccionadas " \
#               " from sga_materiaasignada ma, sga_materia m, sga_asignaturamalla am where ma.status=true and ma.matricula_id=" + str(
#             matricula.id) + " and m.status=true and m.id=ma.materia_id and am.status=true and am.id=m.asignaturamalla_id GROUP by am.nivelmalla_id, am.malla_id order by count(am.nivelmalla_id) desc, am.nivelmalla_id desc limit 1;"
#         cursor.execute(sql)
#         results = cursor.fetchall()
#         nivel = 0
#         for per in results:
#             nivel = per[0]
#             cantidad_seleccionadas = per[1]
#         cantidad_nivel = 0
#
#
#         for asignaturamalla in AsignaturaMalla.objects.filter(nivelmalla__id=nivel, status=True,
#                                                               malla=matricula.inscripcion.mi_malla()):
#             if Materia.objects.filter(nivel__periodo=matricula.nivel.periodo, asignaturamalla=asignaturamalla).exists():
#                 if matricula.inscripcion.estado_asignatura(asignaturamalla.asignatura) != 1:
#                     cantidad_nivel += 1
#
#         porcentaje_seleccionadas = int(round(
#             Decimal((float(cantidad_nivel) * float(PORCIENTO_PERDIDA_PARCIAL_GRATUIDAD)) / 100).quantize(
#                 Decimal('.00')), 0))
#         cobro = 0
#         if matricula.inscripcion.estado_gratuidad == 1 or matricula.inscripcion.estado_gratuidad == 2:
#             if (cantidad_seleccionadas < porcentaje_seleccionadas):
#                 cobro = 1
#             else:
#                 cobro = 2
#         else:
#             if matricula.inscripcion.estado_gratuidad == 2:
#                 cobro = 2
#             else:
#                 cobro = 3
#
#         if matricula.inscripcion.persona.tiene_otro_titulo() or PerdidaGratuidad.objects.filter(
#                 inscripcion=matricula.inscripcion).exists():
#             cobro = 3
#
#         # cobro = 3 # BORRAR SOLO ES PARA PROBAR TITULO - SOLO SE HABILITA PARA PRUEBAS OJOOOO
#
#         if matricula.tiene_pagos_matricula():
#             # cuando tiene pagos de los rubros de la matricula, se le generara un rubro aparta por el valor de la matricula, y el valor adicional de los 10$ pero que no se pase de los 10$
#             matricula.elimina_rubro_matricula_adicional()
#             calculos_finanzas_adicional_aux(matricula, cobro)
#         else:
#             matricula.elimina_rubro_matricula()
#             calculos_finanzas(matricula, cobro)
#         matricula.actualiza_matricula()
#
# def calculos_finanzas_adicional_aux(matricula, cobro):
#     # costo de matricula carlos loyola 03-03-2017
#     from sagest.models import TipoOtroRubro, Rubro
#     persona = matricula.inscripcion.persona
#     periodo = matricula.nivel.periodo
#     if matricula.matriculagruposocioeconomico_set.filter(status=True).exists():
#         valorgrupoeconomico = matricula.matriculagruposocioeconomico_set.filter(status=True)[
#             0].gruposocioeconomico.periodogruposocioeconomico_set.filter(periodo=periodo, status=True)[0].valor
#     else:
#         valorgrupoeconomico = matricula.inscripcion.persona.fichasocioeconomicainec_set.all()[
#             0].grupoeconomico.periodogruposocioeconomico_set.filter(periodo=periodo, status=True)[0].valor
#     porcentaje_gratuidad = periodo.porcentaje_gratuidad
#     valor_maximo = periodo.valor_maximo
#     rubro_anterior_matricula = None
#     if Rubro.objects.filter(matricula=matricula, relacionados__isnull=True, status=True).exists():
#         rubro_anterior_matricula = Rubro.objects.filter(matricula=matricula, relacionados__isnull=True, status=True)[0]
#
#     costo_materia_total = 0
#     tiporubroarancel = TipoOtroRubro.objects.filter(pk=RUBRO_ARANCEL)[0]
#     tiporubromatricula = TipoOtroRubro.objects.filter(pk=RUBRO_MATRICULA)[0]
#
#     if cobro > 0:
#         for materiaasignada in matricula.materiaasignada_set.filter(status=True, retiramateria=False):
#             costo_materia = 0
#             if cobro == 1:
#                 costo_materia = Decimal(Decimal(materiaasignada.materia.creditos).quantize(
#                     Decimal('.01')) * valorgrupoeconomico).quantize(Decimal('.01'))
#             else:
#                 if cobro == 2:
#                     if materiaasignada.matriculas > 1:
#                         costo_materia = Decimal(Decimal(materiaasignada.materia.creditos).quantize(
#                             Decimal('.01')) * valorgrupoeconomico).quantize(Decimal('.01'))
#                 else:
#                     costo_materia = Decimal(Decimal(materiaasignada.materia.creditos).quantize(
#                         Decimal('.01')) * valorgrupoeconomico).quantize(Decimal('.01'))
#             costo_materia_total += costo_materia
#
#     valor_pagado_arancel = 0
#     if Rubro.objects.filter(matricula=matricula, status=True, tipo=tiporubroarancel, cancelado=True).exists():
#         valor_pagado_arancel = \
#         Rubro.objects.filter(matricula=matricula, status=True, tipo=tiporubroarancel, cancelado=True).aggregate(
#             suma=Sum('valor'))['suma']
#         costo_materia_total = costo_materia_total - valor_pagado_arancel
#
#     if costo_materia_total > 0:
#         matricula.estado_matricula = 1
#         matricula.save()
#         valor_porcentaje = Decimal((costo_materia_total * porcentaje_gratuidad) / 100).quantize(Decimal('.01'))
#         rubro = Rubro(tipo=tiporubroarancel,
#                       persona=persona,
#                       relacionados=rubro_anterior_matricula,
#                       matricula=matricula,
#                       # contratorecaudacion = None,
#                       nombre=tiporubroarancel.nombre + ' - ' + periodo.nombre,
#                       cuota=1,
#                       fecha=datetime.now().date(),
#                       fechavence=datetime.now().date() + timedelta(days=1),
#                       valor=costo_materia_total,
#                       iva_id=1,
#                       valoriva=0,
#                       valortotal=costo_materia_total,
#                       saldo=costo_materia_total,
#                       cancelado=False)
#         rubro.save()
#         valor_porcentaje = Decimal(((Rubro.objects.filter(matricula=matricula, status=True,
#                                                           tipo=tiporubroarancel).aggregate(suma=Sum('valor'))[
#             'suma']) * porcentaje_gratuidad) / 100).quantize(Decimal('.01'))
#         valor = valor_porcentaje
#
#         valor_pagado_matricula = 0
#         if Rubro.objects.filter(matricula=matricula, status=True, tipo=tiporubromatricula).exists():
#             valormulta = Decimal((valor_maximo * PORCENTAJE_MULTA) / 100).quantize(Decimal('.01'))
#             valor_pagado_matricula = 0
#             if Rubro.objects.filter(matricula=matricula, status=True, tipo=tiporubromatricula).exclude(
#                     valor=valormulta).exists():
#                 valor_pagado_matricula = \
#                 Rubro.objects.filter(matricula=matricula, status=True, tipo=tiporubromatricula).exclude(
#                     valor=valormulta)[0].valor
#
#         if valor_pagado_matricula < valor:
#             valor = valor - valor_pagado_matricula
#             if (valor + valor_pagado_matricula) > valor_maximo:
#                 valor = valor_maximo - valor_pagado_matricula
#         else:
#             valor = 0
#
#         if valor > 0:
#             rubro1 = Rubro(tipo=tiporubromatricula,
#                            persona=persona,
#                            relacionados=rubro_anterior_matricula,
#                            matricula=matricula,
#                            nombre=tiporubromatricula.nombre + ' - ' + periodo.nombre,
#                            cuota=1,
#                            fecha=datetime.now().date(),
#                            fechavence=datetime.now().date() + timedelta(days=1),
#                            valor=valor,
#                            iva_id=1,
#                            valoriva=0,
#                            valortotal=valor,
#                            saldo=valor,
#                            cancelado=False)
#             rubro1.save()
#
#         if matricula.tipomatricula_id != 1:
#             valor = Decimal((valor_maximo * PORCENTAJE_MULTA) / 100).quantize(Decimal('.01'))
#             if not Rubro.objects.filter(matricula=matricula, status=True, tipo=tiporubromatricula, valor=valor).exists():
#                 rubro1 = Rubro(tipo=tiporubromatricula,
#                                persona=persona,
#                                relacionados=None,
#                                matricula=matricula,
#                                nombre=tiporubromatricula.nombre + ' - ' + periodo.nombre,
#                                cuota=1,
#                                fecha=datetime.now().date(),
#                                fechavence=datetime.now().date() + timedelta(days=1),
#                                valor=valor,
#                                iva_id=1,
#                                valoriva=0,
#                                valortotal=valor,
#                                saldo=valor,
#                                cancelado=False)
#                 rubro1.save()
#     else:
#         if matricula.tipomatricula_id == 1:
#             matricula.estado_matricula = 2
#             matricula.save()
#
# def calculos_finanzas(matricula, cobro):
#     # costo de matricula carlos loyola 03-03-2017
#     from sagest.models import TipoOtroRubro, Rubro
#     from django.db import transaction
#     from django.http import JsonResponse
#     persona = matricula.inscripcion.persona
#     periodo = matricula.nivel.periodo
#     if not matricula.inscripcion.persona.fichasocioeconomicainec_set.all().exists():
#         transaction.set_rollback(True)
#         return JsonResponse({"result": "bad", "reload": False,
#                              "mensaje": u"No puede matricularse, debe llenar la ficha socioeconomica"})
#     valorgrupoeconomico = matricula.inscripcion.persona.fichasocioeconomicainec_set.all()[
#         0].grupoeconomico.periodogruposocioeconomico_set.filter(periodo=periodo, status=True)[0].valor
#     porcentaje_gratuidad = periodo.porcentaje_gratuidad
#     valor_maximo = periodo.valor_maximo
#     costo_materia_total = 0
#     tiporubroarancel = TipoOtroRubro.objects.filter(pk=RUBRO_ARANCEL)[0]
#     tiporubromatricula = TipoOtroRubro.objects.filter(pk=RUBRO_MATRICULA)[0]
#     if cobro > 0:
#         for materiaasignada in matricula.materiaasignada_set.filter(status=True, retiramateria=False):
#             costo_materia = 0
#             if cobro == 1:
#                 costo_materia = Decimal(Decimal(materiaasignada.materia.creditos).quantize(
#                     Decimal('.01')) * valorgrupoeconomico).quantize(Decimal('.01'))
#             else:
#                 if cobro == 2:
#                     if materiaasignada.matriculas > 1:
#                         costo_materia = Decimal(Decimal(materiaasignada.materia.creditos).quantize(
#                             Decimal('.01')) * valorgrupoeconomico).quantize(Decimal('.01'))
#                 else:
#                     costo_materia = Decimal(Decimal(materiaasignada.materia.creditos).quantize(
#                         Decimal('.01')) * valorgrupoeconomico).quantize(Decimal('.01'))
#             costo_materia_total += costo_materia
#
#     if costo_materia_total > 0:
#         valor_porcentaje = Decimal((costo_materia_total * porcentaje_gratuidad) / 100).quantize(Decimal('.01'))
#         rubro = Rubro(tipo=tiporubroarancel,
#                       persona=persona,
#                       relacionados=None,
#                       matricula=matricula,
#                       # contratorecaudacion = None,
#                       nombre=tiporubroarancel.nombre + ' - ' + periodo.nombre,
#                       cuota=1,
#                       fecha=datetime.now().date(),
#                       fechavence=datetime.now().date() + timedelta(days=1),
#                       valor=costo_materia_total,
#                       iva_id=1,
#                       valoriva=0,
#                       valortotal=costo_materia_total,
#                       saldo=costo_materia_total,
#                       cancelado=False)
#         rubro.save()
#
#         valor = valor_porcentaje
#         if valor_porcentaje > valor_maximo:
#             valor = valor_maximo
#         rubro1 = Rubro(tipo=tiporubromatricula,
#                        persona=persona,
#                        relacionados=rubro,
#                        matricula=matricula,
#                        # contratorecaudacion = None,
#                        nombre=tiporubromatricula.nombre + ' - ' + periodo.nombre,
#                        cuota=1,
#                        fecha=datetime.now().date(),
#                        fechavence=datetime.now().date() + timedelta(days=1),
#                        valor=valor,
#                        iva_id=1,
#                        valoriva=0,
#                        valortotal=valor,
#                        saldo=valor,
#                        cancelado=False)
#         rubro1.save()
#
#         # valor multa por no se ordinaria
#         if matricula.tipomatricula_id != 1:
#             valor = Decimal((valor_maximo * PORCENTAJE_MULTA) / 100).quantize(Decimal('.01'))
#             rubro1 = Rubro(tipo=tiporubromatricula,
#                            persona=persona,
#                            relacionados=None,
#                            matricula=matricula,
#                            # contratorecaudacion = None,
#                            nombre=tiporubromatricula.nombre + ' - ' + periodo.nombre,
#                            cuota=1,
#                            fecha=datetime.now().date(),
#                            fechavence=datetime.now().date() + timedelta(days=1),
#                            valor=valor,
#                            iva_id=1,
#                            valoriva=0,
#                            valortotal=valor,
#                            saldo=valor,
#                            cancelado=False)
#             rubro1.save()
#
#     else:
#         if matricula.tipomatricula_id == 1:
#             matricula.estado_matricula = 2
#             matricula.save()
#         else:
#             valor = Decimal((valor_maximo * PORCENTAJE_MULTA) / 100).quantize(Decimal('.01'))
#             rubro1 = Rubro(tipo=tiporubromatricula,
#                            persona=persona,
#                            relacionados=None,
#                            matricula=matricula,
#                            # contratorecaudacion = None,
#                            nombre=tiporubromatricula.nombre + ' - ' + periodo.nombre,
#                            cuota=1,
#                            fecha=datetime.now().date(),
#                            fechavence=datetime.now().date() + timedelta(days=1),
#                            valor=valor,
#                            iva_id=1,
#                            valoriva=0,
#                            valortotal=valor,
#                            saldo=valor,
#                            cancelado=False)
#             rubro1.save()
#             # costo de matricula carlos loyola 03-03-2017
#
# for matricula in Matricula.objects.filter(id__in=[361082,
# 361438,
# 361789,
# 361976,
# 362270,
# 363333,
# 363513,
# 364350,
# 364396,
# 364429,
# 364494,
# 364542,
# 364548,
# 364550,
# 364571,
# 364574,
# 364592,
# 364613,
# 364654,
# 364678,
# 364764,
# 364849,
# 364879,
# 364938,
# 365095,
# 365170,
# 365183,
# 365446,
# 365498,
# 365562,
# 365610,
# 365717,
# 365740,
# 365846,
# 365862,
# 365869,
# 365879,
# 366033,
# 366043,
# 366060,
# 366201,
# 379708,
# 379785,
# 379814,
# 379850,
# 379876,
# 379917,
# 379979,
# 379991,
# 379994,
# 379997,
# 380004,
# 380039,
# 380056,
# 380065,
# 380071,
# 380097,
# 380127,
# 380134,
# 380144,
# 380159,
# 380171,
# 380172,
# 380186,
# 380197,
# 380212,
# 380236,
# 380237,
# 380238,
# 380244,
# 380255,
# 380265,
# 380289,
# 380290,
# 380305,
# 380306,
# 380320,
# 380322,
# 380327,
# 380329,
# 380341,
# 380343]):
#     agregacion_aux(matricula)
#     print(u"%s"%matricula)


#

# admision1@unemi.edu.ec


# historial_proceso_persona = My_HistorialProcesoSubirMatrizInscripcion.objects.get(id=10)
# historial_proceso_persona_observaciones_success = My_HistorialSubirMatrizInscripcion.objects.get(id=8)
#
# observaciones_success = historial_proceso_persona_observaciones_success.observaciones()
# dataErrorObservaciones = []
# dataSuccessObservaciones = []
# errorfichero = False
# sede = Sede.objects.get(pk=1)
# row = 0
# persona=Persona.objects.get(id=20539)
# for arrDataSuccess in observaciones_success[0].observacion:
#     print(arrDataSuccess)
#     row += 1
#     dataFila = {}
#     oCarrera = None
#     oPersona = None
#     oModalidad = None
#     gratuidad = None
#     titulo_tercer = None
#     cupo_aceptado = None
#     segunda_carrera = None
#     for key, value in arrDataSuccess.items():
#         if key in ['CARRERA_ID', 'GRATUIDAD', 'TITULO_TERCER_NIVEL', 'CUPO_ACEPTADO_ACTIVO', '2DA_CARRERA',
#                    'MODALIDAD_ID', 'DISCAPACIDAD', 'RAZA_ID', 'PERSONA_ID']:
#             if key == 'CARRERA_ID':
#                 oCarrera = Carrera.objects.get(pk=value)
#             if key == 'PERSONA_ID':
#                 oPersona = Persona.objects.get(pk=value)
#             if key == 'MODALIDAD_ID':
#                 oModalidad = Modalidad.objects.get(pk=value)
#             if key == 'GRATUIDAD':
#                 gratuidad = value
#             if key == 'TITULO_TERCER_NIVEL':
#                 titulo_tercer = True if value == 'Si' or value == 'True' or value == "1" else False
#             if key == 'CUPO_ACEPTADO_ACTIVO':
#                 cupo_aceptado = True if value == 'Si' or value == 'True' or value == "1" else False
#             if key == '2DA_CARRERA':
#                 segunda_carrera = True if value == 'Si' or value == 'True' or value == "1" else False
#             dataFila[key] = value
#
#     if oCarrera and oPersona and oModalidad and sede:
#         if not Inscripcion.objects.filter(status=True,persona=oPersona, carrera=oCarrera).exists():
#             inscripcion = Inscripcion(persona=oPersona,
#                                       fecha=datetime.now().date(),
#                                       carrera=oCarrera,
#                                       coordinacion=oCarrera.coordinacion_carrera(),
#                                       modalidad=oModalidad,
#                                       sede=sede,
#                                       colegio="N/S" if not oPersona.inscripcion_set.filter().exists() else
#                                       oPersona.inscripcion_set.filter()[0].colegio)
#             inscripcion.save(usuario_id=persona.usuario.id)
#             oPersona.crear_perfil(inscripcion=inscripcion, visible=False)
#             documentos = DocumentosDeInscripcion(inscripcion=inscripcion,
#                                                  titulo=False,
#                                                  acta=False,
#                                                  cedula=False,
#                                                  votacion=False,
#                                                  actaconv=False,
#                                                  partida_nac=False,
#                                                  pre=False,
#                                                  observaciones_pre='',
#                                                  fotos=False)
#             documentos.save(usuario_id=persona.usuario.id)
#             inscripcion.preguntas_inscripcion()
#             inscripcion.malla_inscripcion()
#             inscripcion.actualizar_nivel()
#         else:
#             inscripcion = Inscripcion.objects.filter(persona=oPersona, carrera=oCarrera)[0]
#             mallaalu = inscripcion.inscripcionmalla_set.all()
#             mallaalu.delete()
#             inscripcion.modalidad = oModalidad
#             inscripcion.malla_inscripcion()
#             inscripcion.save(usuario_id=persona.usuario.id)
#         dataFila['INSCRIPCION_ID'] = inscripcion.id
#         print(u"%s"%inscripcion)
#         if not gratuidad:
#             observacion = f"Reportado por la SENESCYT"
#             if titulo_tercer:
#                 observacion = f"{observacion}, Registra TITULO TERCER NIVEL REGISTRO SNIESE"
#             if cupo_aceptado:
#                 observacion = f" {observacion}, Registra CUPO ACEPTADO Y ACTIVO EN SENESCYT"
#             if segunda_carrera:
#                 observacion = f" {observacion}, Registra SOLICITUD DE SEGUNDA CARRERA EN RAES"
#             if not PerdidaGratuidad.objects.values("id").filter(status=True, inscripcion=inscripcion).exists():
#                 ePerdidaGratuidad = PerdidaGratuidad(inscripcion=inscripcion,
#                                                      motivo=1,
#                                                      titulo=None,
#                                                      titulo_sniese=titulo_tercer,
#                                                      cupo_aceptado_senescyt=cupo_aceptado,
#                                                      segunda_carrera_raes=segunda_carrera,
#                                                      observacion=observacion
#                                                      )
#             else:
#                 ePerdidaGratuidad = PerdidaGratuidad.objects.filter(status=True, inscripcion=inscripcion)[0]
#                 ePerdidaGratuidad.motivo = 1
#                 ePerdidaGratuidad.titulo = None
#                 ePerdidaGratuidad.titulo_sniese = titulo_tercer
#                 ePerdidaGratuidad.cupo_aceptado_senescyt = cupo_aceptado
#                 ePerdidaGratuidad.segunda_carrera_raes = segunda_carrera
#                 ePerdidaGratuidad.observacion = observacion
#
#             ePerdidaGratuidad.save(usuario_id=persona.usuario.id)
#     else:
#         dataErrorObservaciones.append({'fila': row,
#                                        'columna': 'Persona',
#                                        'valor': dataFila['PERSONA_ID'],
#                                        'mensaje': 'Error, no se puede procesar la persona'
#                                        })
#         print(dataErrorObservaciones)
#
#     if dataFila:
#         dataSuccessObservaciones.append(dataFila)
# print("PROCESO DE CREAR INSCRIPCIÓN")

#
# persona_ids = Matricula.objects.filter(status=True, nivel__periodo_id=119, nivelmalla_id=1).values('inscripcion__persona_id').distinct()
# persona_ids_aux = Matricula.objects.filter(status=True, nivel__periodo_id=136, materiaasignada__isnull=False).values('inscripcion__persona_id').distinct()
# personas_ids = persona_ids | persona_ids_aux
# personas_ids = persona_ids_aux
# MATRICULAS DEL PERIODO JUNIO A SEPTIEMBRE 2021 EN ADMISION
# repetidores = Matricula.objects.filter(id=337239,status=True,  aprobado=False, nivel__periodo_id=123).exclude(inscripcion__persona_id__in=personas_ids)
# cont=0
# for repe in repetidores:
#     # if repe.materiaasignada_set.filter(estado_id=2).count()>0:
#     if repe.mismaterias().count() == 3:
#             cont+=1
#
# print(cont)


# matriz = My_SubirMatrizInscripcion.objects.get(pk=9)
# periodo=Periodo.objects.get(id=136)
# persona=Persona.objects.get(id=20539)
# vrProcess = matriz.matriculacion_senescyt(periodo, persona)


# workbook = xlrd.open_workbook("matricula_admision.xlsx")
# sheet = workbook.sheet_by_index(0)
# cedula = 0
# carrera = 0
# periodo = Periodo.objects.get(pk=136)
# linea=0
# import time
# try:
#     for rowx in range(sheet.nrows):
#         if linea>18999 and linea<20000:
#             cols = sheet.row_values(rowx)
#             cedula = cols[0].strip().upper()
#             persona=None
#             idcarrera=None
#             if Persona.objects.filter(Q(cedula=cedula) | Q(pasaporte=cedula), status=True).exists():
#                 datospersona = Persona.objects.filter(Q(cedula=cedula) | Q(pasaporte=cedula), status=True)[0]
#                 lista = []
#                 if datospersona.email:
#                     lista.append(datospersona.email)
#                     if linea <=10:
#                         lista.append("jplacesc@unemi.edu.ec")
#                     send_html_mail(u"[IMPORTANTE] UNEMI - Credenciales de acceso al Sistema de Gestión Académica",
#                                    "emails/email_notificacion_credenciales_admision.html",
#                                    {
#                                        'sistema': u'SGA - UNEMI',
#                                        'fecha': datetime.now().date(),
#                                        'hora': datetime.now().time(),
#                                        'persona': datospersona,
#                                        't': miinstitucion()
#                                    },
#                                    lista,
#                                    [],
#                                    cuenta=variable_valor('CUENTAS_CORREOS')[35]
#                                    )
#                     print("%s - Correo enviado %s +++++ %s " % (linea,datospersona,datospersona.email))
#                     time.sleep(2)
#                 else:
#                     print("No tiene correo %s +++++  " % (datospersona))
#         linea += 1
# except Exception as ex:
#     print(ex)
# matriz = My_SubirMatrizInscripcion.objects.get(pk=9)
# periodoactual=Periodo.objects.get(id=136)
# periodoanterior=Periodo.objects.get(id=123)
# persona=Persona.objects.get(id=20539)
# vrProcess = matriz.matricula_2_repetidores_admision(periodoactual,periodoanterior, persona)

#
# matriculas_admision_1s2021=Matricula.objects.filter(status=False,nivel__periodo_id=123,
#                                                                       termino=False,automatriculaadmision=True,
#                                                                       aprobado=False)
# print(matriculas_admision_1s2021.count())
# for mat in matriculas_admision_1s2021:
#     mat.delete()

# # matriculas_pregrado=Matricula.objects.filter(status=True,nivel__periodo_id=113, termino=False,automatriculapregrado=True)
# # print(u"%s"%matriculas_pregrado.count())
# # con = 0
# # for mat in matriculas_pregrado:
# #     mat.delete()
# #     con += 1
# #     print(u"%s.- %s" % (con, mat))


# try:
#     persona=Persona.objects.get(id=20539)
#     periodoactual=Periodo.objects.get(id=136)
#     linea=0
#     # for matricula in Matricula.objects.filter(id__in=[422821]):
#
#     for matricula in Matricula.objects.filter(id__in=[423181]):
#
#         materias_asignadas=matricula.materiaasignada_set.filter(status=True)
#         if materias_asignadas:
#             for materiaasignada in materias_asignadas:
#                 materiaasignada.cobroperdidagratuidad = True
#                 materiaasignada.save(usuario_id=persona.usuario.id)
#                 materiaasignada.asistencias()
#                 materiaasignada.evaluacion()
#                 materiaasignada.mis_planificaciones()
#                 materiaasignada.save(usuario_id=persona.usuario.id)
#                 print(materiaasignada)
#             matricula.actualizar_horas_creditos()
#             matricula.estado_matricula = 2
#             matricula.save(usuario_id=persona.usuario.id)
#             matricula.calcula_nivel()
#             matricula.inscripcion.actualizar_nivel()
#
#             if matricula.inscripcion.sesion_id == 13:
#                 tiporubromatricula = TipoOtroRubro.objects.get(pk=3019)
#             else:
#                 tiporubromatricula = TipoOtroRubro.objects.get(pk=3011)
#
#             num_materias = MateriaAsignada.objects.filter(matricula=matricula, cobroperdidagratuidad=True).exclude(materia__asignatura_id=4837).count()
#             if num_materias > 0:
#                 valor_x_materia = 15
#                 valor_total = num_materias * valor_x_materia
#                 matricula.estado_matricula = 1
#                 matricula.save(usuario_id=persona.usuario.id)
#                 if not Rubro.objects.filter(persona=matricula.inscripcion.persona, matricula=matricula).exists():
#                     rubro1 = Rubro(tipo=tiporubromatricula,
#                                    persona=matricula.inscripcion.persona,
#                                    matricula=matricula,
#                                    nombre=tiporubromatricula.nombre + ' - ' + periodoactual.nombre,
#                                    cuota=1,
#                                    fecha=datetime.now().date(),
#                                    fechavence=datetime.now().date() + timedelta(days=6),
#                                    valor=valor_total,
#                                    iva_id=1,
#                                    valoriva=0,
#                                    valortotal=valor_total,
#                                    saldo=valor_total,
#                                    cancelado=False)
#                     rubro1.save(usuario_id=persona.usuario.id)
#                     print(rubro1)
#         else:
#             print(u"No existen materias asignadas - %s"%matricula)
#         linea += 1
# except Exception as ex:
#     print(ex)


#
# for materia in  Materia.objects.filter(id__in=[39805, 39789, 39791, 43039, 43042, 41146, 41067, 39499, 42941, 38870, 45123, 42936, 38868, 38679, 38936, 38689, 38939, 38674, 38656, 38909, 44589, 38980, 43222, 38972,44593,
# 38969,  44597, 38970, 38973, 44594, 38971, 38991, 38230, 43335, 38204, 38205, 38210, 38208, 38212,43317,38206,38211,43316,38209,38207,43324,38232,38114,38101,38090,38089,38086,
# 38104,38105,38110,38118,38102,38120,38111,38106,38087,38099,38113,38112,38116,38107,38119,38117,38097,38108,40257,40266,44619,44730,44742,44624,44664,
# 44722,44755,44763,44620,44660,43051,39681]):
# materia=Materia.objects.get(id=41075 )
# materiasasignadas = MateriaAsignada.objects.filter(id=1578417)
# grupoprofesor=None
# if GruposProfesorMateria.objects.filter(status=True, profesormateria__materia=materia).exists():
#     grupoprofesor = GruposProfesorMateria.objects.get(status=True, profesormateria__materia=materia)
#     print(grupoprofesor)
# else:
#     print(materia)
# if grupoprofesor:
#     for materiaaasignada in materiasasignadas:
#         try:
#             from itertools import chain
#             listaprofemateriaid_congrupo = []
#             listagrupos = []
#             listamateriaasignada = []
#             if AlumnosPracticaMateria.objects.values('id').filter(materiaasignada=materiaaasignada,
#                                                                   status=True).exists():
#                 alumnopractica =  AlumnosPracticaMateria.objects.filter(materiaasignada=materiaaasignada, status=True)[0]
#                 alumnopractica.profesormateria = grupoprofesor.profesormateria
#                 alumnopractica.grupoprofesor = grupoprofesor
#             else:
#                 alumnopractica = AlumnosPracticaMateria(profesormateria=grupoprofesor.profesormateria,
#                                                         materiaasignada=materiaaasignada,
#                                                         grupoprofesor=grupoprofesor)
#             alumnopractica.save()
#             print(alumnopractica)
#         except Exception as ex:
#             print( u'Error al matricular estudiante. %s'%materiaaasignada)
#

# poner asistencia dia lunes


# fecha = (datetime(2021, 11, 29, 0, 0, 0)).date()
# print("FECHA A PROCESAR: " + fecha.__str__() + "\r")
# for cl in Clase.objects.filter(activo=True, inicio__lte=fecha, fin__gte=fecha, dia=(fecha.weekday() + 1), status=True, materia__nivel__periodo_id=119):
#     if cl.materia.profesor_principal():
#         if LeccionGrupo.objects.filter(profesor=cl.materia.profesor_principal(), turno=cl.turno, fecha=fecha).exists():
#             lecciongrupo = LeccionGrupo.objects.get(profesor=cl.materia.profesor_principal(), turno=cl.turno, fecha=fecha)
#         else:
#             lecciongrupo = LeccionGrupo(profesor=cl.materia.profesor_principal(),
#                                         turno=cl.turno,
#                                         aula=cl.aula,
#                                         dia=cl.dia,
#                                         fecha=fecha,
#                                         horaentrada=cl.turno.comienza,
#                                         horasalida=cl.turno.termina,
#                                         abierta=False,
#                                         automatica=True,
#                                         contenido='REGISTRO MASIVO 2021 - AUTORIZADO POR DIRECTOR TICS',
#                                         observaciones='REGISTRO MASIVO 2021 - AUTORIZADO POR DIRECTOR TICS')
#             lecciongrupo.save()
#         if Leccion.objects.filter(clase=cl, fecha=fecha).exists():
#             leccion = Leccion.objects.get(clase=cl, fecha=fecha)
#         else:
#             leccion = Leccion(clase=cl,
#                               fecha=fecha,
#                               horaentrada=cl.turno.comienza,
#                               horasalida=cl.turno.termina,
#                               abierta=True,
#                               contenido=lecciongrupo.contenido,
#                               observaciones=lecciongrupo.observaciones)
#             leccion.save()
#         if not lecciongrupo.lecciones.filter(pk=leccion.id).exists():
#             lecciongrupo.lecciones.add(leccion)
#         if AsistenciaLeccion.objects.filter(leccion=leccion).exists():
#             for asis in AsistenciaLeccion.objects.filter(leccion=leccion):
#                 if not asis.asistio:
#                     asis.asistio = True
#                     asis.save()
#                     mateasig = asis.materiaasignada
#                     mateasig.save(actualiza=True)
#                     mateasig.actualiza_estado()
#         else:
#             for materiaasignada in cl.materia.asignados_a_esta_materia():
#                 if not AsistenciaLeccion.objects.filter(leccion=leccion, materiaasignada=materiaasignada).exists():
#                     asistencialeccion = AsistenciaLeccion(leccion=leccion,
#                                                           materiaasignada=materiaasignada,
#                                                           asistio=True)
#                     asistencialeccion.save()
#                     materiaasignada.save(actualiza=True)
#                     materiaasignada.actualiza_estado()
#                 # guardar temas de silabo
#         lecciongrupo.save()
#         print(cl)
#

# fin poner asistencia dia lunes



# print ("What is your name?")
# student_name= input()
# homework= float(input())
# classwork= float(input())
# test= float(input())
# average= (homework+ classwork + test)/3
# print(student_name, "your score for this partial is: %s" % average)

# for video in VideoMagistralSilaboSemanal.objects.filter(id__in=[4206,4207,4208,4209,4210,4211,4212,4213,4214]):
#     print(u"%s"%video)
#     video.delete()

# for tarea in TareaSilaboSemanal.objects.filter(id__in=[36846,36845]):
#     print(u"%s"%tarea)
#     tarea.delete()
#
# for tarea in TareaSilaboSemanal.objects.filter(id__in=[38995,39006]):
#     print(u"%s"%tarea)
#     tarea.delete()

# for materia in Materia.objects.filter(id__in=[45698]):
#     #################################################################################################################
#     # AGREGAR ESTUDIANTE
#     #################################################################################################################
#     from sga.funciones import log
#
#     if materia.idcursomoodle:
#         contador = 0
#         cursoid = materia.idcursomoodle
#         for estudiante in MateriaAsignada.objects.filter(id=1760740):
#             try:
#                 contador += 1
#                 bandera = 0
#                 persona = estudiante.matricula.inscripcion.persona
#                 username = persona.usuario.username
#                 bestudiante = moodle.BuscarUsuario(materia.nivel.periodo, 2, 'username', username)
#                 estudianteid = 0
#                 if not bestudiante:
#                     bestudiante = moodle.BuscarUsuario(materia.nivel.periodo, 2, 'username', username)
#
#                 if bestudiante['users']:
#                     if 'id' in bestudiante['users'][0]:
#                         estudianteid = bestudiante['users'][0]['id']
#                 else:
#                     bandera = 1
#                     idnumber_user = persona.identificacion()
#                     notuser = moodle.BuscarUsuario(materia.nivel.periodo, 2, 'username', idnumber_user)
#                     if not notuser:
#                         notuser = moodle.BuscarUsuario(materia.nivel.periodo, 2, 'username', idnumber_user)
#                     if notuser['users']:
#                         elminar = moodle.EliminarUsuario(materia.nivel.periodo, 2, notuser['users'][0]['id'])
#
#                     bestudiante = moodle.CrearUsuario(materia.nivel.periodo, 2, u'%s' % persona.usuario.username,
#                                                       u'%s' % persona.identificacion(),
#                                                       u'%s' % persona.nombres,
#                                                       u'%s %s' % (persona.apellido1, persona.apellido2),
#                                                       u'%s' % persona.emailinst,
#                                                       idnumber_user,
#                                                       u'MILAGRO',
#                                                       u'ECUADOR')
#                     estudianteid = bestudiante[0]['id']
#
#                 cursor = connections['db_moodle_virtual'].cursor()
#                 sql = """ UPDATE mooc_user
#                                     SET idnumber ='%s'
#                                     WHERE username='%s'
#                                     """ % (persona.identificacion(), persona.usuario.username)
#                 cursor.execute(sql)
#
#                 if estudianteid > 0:
#                     rolest = moodle.EnrolarCurso(materia.nivel.periodo, 2, materia.nivel.periodo.rolestudiante,
#                                                  estudianteid, cursoid)
#                     if persona.idusermoodle != estudianteid:
#                         persona.idusermoodle = estudianteid
#                         persona.save()
#                 print('************Estudiante: %s *** %s' % (contador, persona))
#             except Exception as ex:
#                 from sagest.models import ControlMatriculaMoodle
#                 control = ControlMatriculaMoodle(persona=persona, materia=estudiante.materia,
#                                                  materiaasignada=estudiante,
#                                                  error=u"%s" % ex )
#                 control.save()
#                 log(u'Moodle Error al crear Estudiante: %s %s (%s)' % (persona, estudiante, estudiante.id), None, "add", User.objects.get(pk=1))
#                 print('Error al crear estudiante %s' % ex)
#         materia.quitar_estudiantes_curso(moodle, 2)

# for test in TestSilaboSemanal.objects.filter(id__in=[23925]):
#     print(u"%s"%test)
#     test.delete()
#
# for tarea in TareaSilaboSemanal.objects.filter(id__in=[37989]):
#     print(u"%s"%tarea)
#     tarea.delete()

# print("Inicia")
# for materiaasignada in MateriaAsignada.objects.filter(id__in=[1816429,         1816756]):
#     materiaasignada.delete()
#     print("Elimina")
# print("Finaliza")
# INICIO PROCESO MATRICULA ESTUDIANTE INGLES


# REPORTE DE POBLACIONA  AMTRICULAR

# import xlwt
# from xlwt import *
# from django.http import HttpResponse
#
# response = HttpResponse(content_type="application/ms-excel")
# response['Content-Disposition'] = 'attachment; filename=reporte_moulos_ingles.xls'
# style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
# style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
# style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
# title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
# style1 = easyxf(num_format_str='D-MMM-YY')
# font_style = XFStyle()
# font_style.font.bold = True
# font_style2 = XFStyle()
# font_style2.font.bold = False
# wb = xlwt.Workbook()
# ws = wb.add_sheet('Sheetname')
# estilo = xlwt.easyxf('font: height 350, name Arial, colour_index black, bold on, italic on; align: wrap on, vert centre, horiz center;')
# ws.write_merge(0, 0, 0, 9, 'UNIVERSIDAD ESTATAL DE MILAGRO', estilo)
# output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
# nombre = "Lista" + datetime.now().strftime('%Y%m%d_%H%M%S') + ".xls"
# filename = os.path.join(output_folder, nombre)
# columns = [(u"Id", 6000),
#            (u"carrera", 6000),
#            (u"cedula", 6000),
#            (u"alumno", 6000),
#            (u"modulo", 6000),
#            (u"cant matricula", 6000),
#            (u"perdida de gratuidad", 6000),
#            ]
# row_num = 3
# for col_num in range(len(columns)):
#     ws.write(row_num, col_num, columns[col_num][0], font_style)
#     ws.col(col_num).width = columns[col_num][1]
# row_num = 4
# contador=1
# nivel=Nivel.objects.get(id=658)
# idmatriculas=Matricula.objects.values_list('id',flat=True).filter(status=True, nivel__periodo_id=119,retiradomatricula=False,materiaasignada__materia__id__in=[46031,46032,46033,46034,46035,46036,46037,46038,46039,46040]).distinct()
# for matricula in Matricula.objects.filter(status=True, nivel__periodo_id=119,retiradomatricula=False,
#                                           ).exclude(inscripcion__carrera__id__in=[7,138,129,90,157]
#                                                     ).exclude(id__in=idmatriculas).distinct():
#     if matricula.inscripcion.ultimo_modulo_ingles_pendiente():
#         modulomalla=matricula.inscripcion.ultimo_modulo_ingles_pendiente()
#         matriculas = matricula.inscripcion.historicorecordacademico_set.filter(asignatura=modulomalla.asignatura,
#                                                                                fecha__lt=nivel.fin).count() + 1
#         ws.write(row_num, 0, u'%s' % matricula.id, font_style2)
#         ws.write(row_num, 1, u'%s' % matricula.inscripcion.carrera, font_style2)
#         ws.write(row_num, 2, u'%s' % matricula.inscripcion.persona.identificacion(), font_style2)
#         ws.write(row_num, 3, u'%s' % matricula.inscripcion.persona, font_style2)
#         ws.write(row_num, 4, u'%s' % modulomalla, font_style2)
#         ws.write(row_num, 5, u'%s' % matriculas, font_style2)
#         ws.write(row_num, 6, u'%s' % "SI" if matricula.inscripcion.persona.tiene_otro_titulo(matricula.inscripcion) else "NO", font_style2)
#         print(u"%s - %s" % (contador,matricula))
#         row_num += 1
#         contador += 1
# wb.save(filename)
# print("FIN: ", filename)



# PROCESO DE MATRICULA INGLÉS ESTUDIANTE

# import xlwt
# from xlwt import *
# from django.http import HttpResponse
#
# response = HttpResponse(content_type="application/ms-excel")
# response['Content-Disposition'] = 'attachment; filename=reporte_moulos_ingles.xls'
# style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
# style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
# style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
# title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
# style1 = easyxf(num_format_str='D-MMM-YY')
# font_style = XFStyle()
# font_style.font.bold = True
# font_style2 = XFStyle()
# font_style2.font.bold = False
# wb = xlwt.Workbook()
# ws = wb.add_sheet('Sheetname')
# estilo = xlwt.easyxf('font: height 350, name Arial, colour_index black, bold on, italic on; align: wrap on, vert centre, horiz center;')
# ws.write_merge(0, 0, 0, 9, 'UNIVERSIDAD ESTATAL DE MILAGRO', estilo)
# output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
# nombre = "Lista" + datetime.now().strftime('%Y%m%d_%H%M%S') + ".xls"
# filename = os.path.join(output_folder, nombre)
# columns = [(u"Id", 6000),
#            (u"carrera", 6000),
#            (u"cedula", 6000),
#            (u"alumno", 6000),
#            (u"modulo", 6000),
#            (u"cant matricula", 6000),
#            (u"materia asignada", 6000),
#            (u"perdida de gratuidad", 6000),
#            ]
# row_num = 3
# for col_num in range(len(columns)):
#     ws.write(row_num, col_num, columns[col_num][0], font_style)
#     ws.col(col_num).width = columns[col_num][1]
# row_num = 4
#
# contador=1
# persona_responsable=Persona.objects.get(id=20539)
# for matricula in Matricula.objects.filter(status=True, nivel__periodo_id=119,retiradomatricula=False, id__in=[423485,423572,423508,398557,382327,396928,395824,423564,398765,
#                                                                                                               422963,398842,374060,385438,423466,381311,380524,379308,383136,378944,
#                                                                                                               398860,375717,398821,374440,392649,383279,423584,394097,375975,378393,
#                                                                                                               423810,423879,423585,423451,423588,423457,398772,398778,398735,423578,
#                                                                                                               398060,394629,390074,423482,423573,423632,423491,393198,377836,376495,
#                                                                                                               377152,378388,372222,376439,377069,379587,377673,370097,394608,380093,
#                                                                                                               377776,391259,373699,375567,398876,398822,393753,376373,389272,376225,
#                                                                                                               376848,377505,368829,398461,379358,389715,378407,423450,381151,391000,
#                                                                                                               393195,387045,378075,377854,423474,397610,423519,383257,390025,379516,
#                                                                                                               372169,380295,423486,377515,378848,392046,396362,378644,423471,378677,
#                                                                                                               395449] ).exclude(inscripcion__carrera__id__in=[7,138,134,129,90,157]):
#     try:
#         bandera=0
#         modulomalla=None
#         ws.write(row_num, 0, u'%s' % matricula.id, font_style2)
#         ws.write(row_num, 1, u'%s' % matricula.inscripcion.carrera, font_style2)
#         ws.write(row_num, 2, u'%s' % matricula.inscripcion.persona.identificacion(), font_style2)
#         ws.write(row_num, 3, u'%s' % matricula.inscripcion.persona, font_style2)
#         id_modulos_ingles_todos=None
#         if matricula.inscripcion.modulos_ingles_mi_malla():
#             id_modulos_ingles_todos=matricula.inscripcion.modulos_ingles_mi_malla().values_list('asignatura_id',flat=True).distinct()
#         if id_modulos_ingles_todos:
#             if matricula.inscripcion.ultimo_modulo_ingles_pendiente():
#                 modulomalla=matricula.inscripcion.ultimo_modulo_ingles_pendiente()
#                 materias=Materia.objects.filter(status=True, nivel_id=658, asignatura=modulomalla.asignatura,inglesepunemi=True )
#                 for materia in materias:
#                     if materia.tiene_capacidad():
#                         matriculas=1
#                         materiaasignada=None
#                         matriculas = matricula.inscripcion.historicorecordacademico_set.filter(asignatura=materia.asignatura).count() + 1
#                         if not MateriaAsignada.objects.filter(matricula=matricula, materia=materia, materia__asignatura=modulomalla.asignatura).exists():
#                             materiaasignada = MateriaAsignada(matricula=matricula,
#                                                               materia=materia,
#                                                               notafinal=0,
#                                                               asistenciafinal=0,
#                                                               cerrado=False,
#                                                               matriculas=matriculas,
#                                                               observaciones='',
#                                                               estado_id=NOTA_ESTADO_EN_CURSO,
#                                                               automatricula=False,
#                                                               importa_nota=False)
#                             materiaasignada.save()
#                             materiaasignada.evaluacion()
#                             creditos = materiaasignada.materia.creditos
#                             if materiaasignada.existe_modulo_en_malla():
#                                 creditos = materiaasignada.materia_modulo_malla().creditos
#                             registro = AgregacionEliminacionMaterias(matricula=matricula,
#                                                                      agregacion=True,
#                                                                      asignatura=materiaasignada.materia.asignatura,
#                                                                      responsable=persona_responsable,
#                                                                      fecha=datetime.now().date(),
#                                                                      creditos=creditos,
#                                                                      nivelmalla=materiaasignada.materia.nivel.nivelmalla if materiaasignada.materia.nivel.nivelmalla else None,
#                                                                      matriculas=materiaasignada.matriculas)
#                             registro.save()
#                             if materiaasignada.matriculas>1 or matricula.inscripcion.persona.tiene_otro_titulo(matricula.inscripcion):
#                                 matricula.calculo_matricula_ingles(materiaasignada)
#                             ws.write(row_num, 4, u'%s' % modulomalla, font_style2)
#                             ws.write(row_num, 5, u'%s' % matriculas, font_style2)
#                             ws.write(row_num, 6, u'%s' % materiaasignada, font_style2)
#                             ws.write(row_num, 7, u'%s' % "SI" if matricula.inscripcion.persona.tiene_otro_titulo(matricula.inscripcion) else "NO", font_style2)
#                             row_num += 1
#                             print(u"(%s) -- %s [%s]" % (contador, materiaasignada, materiaasignada.matricula_id))
#                             contador += 1
#                             bandera=1
#                             break
#                         else:
#                             ws.write(row_num, 4, u'YA TIENE MATRICULA EN EL MODULO' %materia, font_style2)
#                             ws.write(row_num, 5, u'YA TIENE MATRICULA EN EL MODULO' %materia, font_style2)
#                             ws.write(row_num, 6, u'YA TIENE MATRICULA EN EL MODULO' %materia, font_style2)
#                             ws.write(row_num, 7, u'YA TIENE MATRICULA EN EL MODULO' %materia, font_style2)
#                             row_num += 1
#                             print(u"(%s) -- YA TIENE MATRICULA EN EL MODULO %s" % (contador,materia))
#                             contador += 1
#                             bandera = 1
#             else:
#                 ws.write(row_num, 4, u'NO AY ULTIMO MÓDULO DE INGLÉS PENDIENTE', font_style2)
#                 ws.write(row_num, 5, u'NO AY ULTIMO MÓDULO DE INGLÉS PENDIENTE', font_style2)
#                 ws.write(row_num, 6, u'NO AY ULTIMO MÓDULO DE INGLÉS PENDIENTE', font_style2)
#                 ws.write(row_num, 7, u'NO AY ULTIMO MÓDULO DE INGLÉS PENDIENTE', font_style2)
#                 row_num += 1
#                 print(u"(%s) -- NO AY ULTIMO MÓDULO DE INGLÉS PENDIENTE %s" % (contador,matricula))
#                 contador += 1
#                 bandera = 1
#             # else:
#             #     ws.write(row_num, 4, u'YA TIENE MATRICULA EN AL MENOS UN MODULO %s' %matricula, font_style2)
#             #     ws.write(row_num, 5, u'YA TIENE MATRICULA EN AL MENOS UN MODULO %s' %matricula, font_style2)
#             #     ws.write(row_num, 6, u'YA TIENE MATRICULA EN AL MENOS UN MODULO %s' %matricula, font_style2)
#             #     ws.write(row_num, 7, u'YA TIENE MATRICULA EN AL MENOS UN MODULO %s' %matricula, font_style2)
#             #     row_num += 1
#             #     print(u"(%s) -- YA TIENE MATRICULA EN AL MENOS UN MODULO %s" %(matricula, contador))
#             #     contador += 1
#             #     bandera = 1
#         else:
#             ws.write(row_num, 4, u'NO HAY MODULOS DE INGLES EN LA MALLA', font_style2)
#             ws.write(row_num, 5, u'NO HAY MODULOS DE INGLES EN LA MALLA', font_style2)
#             ws.write(row_num, 6, u'NO HAY MODULOS DE INGLES EN LA MALLA', font_style2)
#             ws.write(row_num, 7, u'NO HAY MODULOS DE INGLES EN LA MALLA', font_style2)
#             row_num += 1
#             print(u"(%s) -- NO HAY MODULOS DE INGLES EN LA MALLA" % contador)
#             contador += 1
#             bandera = 1
#         if bandera == 0:
#             ws.write(row_num, 4, u'NO SE MATRICULA - NO HAY CUPO %s -- %s '%(modulomalla,matricula), font_style2)
#             ws.write(row_num, 5, u'NO SE MATRICULA - NO HAY CUPO %s -- %s '%(modulomalla,matricula), font_style2)
#             ws.write(row_num, 6, u'NO SE MATRICULA - NO HAY CUPO %s -- %s '%(modulomalla,matricula), font_style2)
#             ws.write(row_num, 7, u'NO SE MATRICULA - NO HAY CUPO %s -- %s '%(modulomalla,matricula), font_style2)
#             row_num += 1
#             print(u"NO SE MATRICULA - NO HAY CUPO %s -- %s" % (modulomalla,matricula))
#             contador += 1
#     except Exception as ex:
#         ws.write(row_num, 4, u'SALE ERROR %s' % ( matricula), font_style2)
#         ws.write(row_num, 5, u'SALE ERROR %s' % ( matricula), font_style2)
#         ws.write(row_num, 6, u'SALE ERROR %s' % ( matricula), font_style2)
#         ws.write(row_num, 7, u'SALE ERROR %s' % ( matricula), font_style2)
#         row_num += 1
#         contador += 1
#         print('error: %s %s' % (ex,matricula))
#         pass
# wb.save(filename)
# print("FIN: ", filename)




# try:
#     miarchivo = openpyxl.load_workbook("EXAMENES TODOS.xlsx")
#     lista = miarchivo.get_sheet_by_name('TODOS')
#     totallista = lista.rows
#     a=0
#     periodo=Periodo.objects.get(id=119)
#     for filas in totallista:
#         a += 1
#         if a > 2:
#             id_materia = int(filas[6].value)
#             materia=Materia.objects.get(id=id_materia)
#             fecha=convertir_fecha_invertida(filas[13].value)
#             hora_inicio=convertir_hora(filas[14].value)
#             hora_fin=convertir_hora(filas[15].value)
#             caant_alumn=int(filas[16].value)
#             horarioexamen=None
#             if not HorarioExamen.objects.filter(materia=materia,
#                                           detallemodelo_id=37,
#                                           turno_id=1).exists():
#                 horarioexamen = HorarioExamen(materia=materia,
#                                               detallemodelo_id=37,
#                                               fecha=fecha,
#                                               turno_id=1)
#                 horarioexamen.save()
#             else:
#                 horarioexamen=HorarioExamen.objects.filter(materia=materia,
#                                              detallemodelo_id=37,
#                                              turno_id=1)[0]
#                 horarioexamen.fecha=fecha
#                 horarioexamen.save()
#
#             if not HorarioExamenDetalle.objects.filter(horarioexamen=horarioexamen).exists():
#                 detallexamen = HorarioExamenDetalle(horarioexamen=horarioexamen,
#                                                       horainicio=hora_inicio,
#                                                       horafin=hora_fin,
#                                                       cantalumnos=caant_alumn)
#                 detallexamen.save()
#             else:
#                 detallexamen=HorarioExamenDetalle.objects.filter(horarioexamen=horarioexamen)[0]
#                 detallexamen.horainicio=hora_inicio
#                 detallexamen.horafin=hora_fin
#                 detallexamen.cantalumnos=caant_alumn
#                 detallexamen.save()
#             print('Registro ingresado: %s' %materia )
# except Exception as ex:
#         print('error: %s' % ex)





# try:
#     miarchivo = openpyxl.load_workbook("docentes zoom.xlsx")
#     lista = miarchivo.get_sheet_by_name('docentes con licencia')
#     totallista = lista.rows
#     a=0
#     for filas in totallista:
#         a += 1
#         if a > 1:
#             correo=str(filas[0].value)
#             if filas[3].value:
#                 idzoom=filas[3].value.replace(' ','')
#                 if Profesor.objects.filter(status=True,persona__emailinst=correo,urlzoom__icontains=idzoom).exists():
#                     filas[4].value = "COINCIDE ID ZOOM SGA"
#                 else:
#                     filas[4].value = "INCONSISTENCIA ID ZOOM SGA"
#     miarchivo.save("docentes zoom.xlsx")
# except Exception as ex:
#         print('error: %s' % ex)



def verificar_enrolado_moodle(periodo,cursor):
    response = HttpResponse(content_type="application/ms-excel")
    response['Content-Disposition'] = 'attachment; filename=reporteenrolados_salud.xls'
    style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
    style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
    style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
    title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
    style1 = easyxf(num_format_str='D-MMM-YY')
    font_style = XFStyle()
    font_style.font.bold = True
    font_style2 = XFStyle()
    font_style2.font.bold = False
    wb = xlwt.Workbook()
    ws = wb.add_sheet('Sheetname')
    estilo = xlwt.easyxf('font: height 350, name Arial, colour_index black, bold on, italic on; align: wrap on, vert centre, horiz center;')
    ws.write_merge(0, 0, 0, 9, 'UNIVERSIDAD ESTATAL DE MILAGRO', estilo)
    output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
    nombre = "Lista" + datetime.now().strftime('%Y%m%d_%H%M%S') + ".xls"
    filename = os.path.join(output_folder, nombre)
    columns = [(u"Carrera", 6000),
               (u"Sesion", 6000),
               (u"Nivel", 6000),
               (u"id materia", 6000),
               (u"Materia", 6000),
               (u"Paralelo", 6000),
               (u"Estudiante", 6000),
               (u"cedula", 6000),
               (u"Enrolado moodle", 6000),
               (u"Tiene deuda", 6000),
               ]
    row_num = 3
    for col_num in range(len(columns)):
        ws.write(row_num, col_num, columns[col_num][0], font_style)
        ws.col(col_num).width = columns[col_num][1]
    row_num = 4
    materias=Materia.objects.filter(status=True, nivel__periodo=periodo)
    for materia in materias:
        for materiaasignada in materia.materiaasignada_set.filter(status=True,matricula__status=True,retiramateria=False ):
            enrolado="SI"
            tienedeuda="NO"
            query = """ SELECT DISTINCT asi.userid, asi.roleid
                                FROM  mooc_role_assignments asi
                                INNER JOIN MOOC_CONTEXT CON ON asi.CONTEXTID=CON.ID AND ASI.ROLEID=%s
                                INNER JOIN mooc_user usr1 on usr1.id=asi.userid
                                AND CON.INSTANCEID=%s AND usr1.username ='%s'
                        """ % (materia.nivel.periodo.rolestudiante, materia.idcursomoodle,
                               materiaasignada.matricula.inscripcion.persona.usuario.username)
            cursor.execute(query)
            row = cursor.fetchall()
            if not row:
                enrolado = "NO"
                if Rubro.objects.filter(status=True,matricula=materiaasignada.matricula).exists():
                    tienedeuda="SI"
                ws.write(row_num, 0, u'%s' % materia.asignaturamalla.malla.carrera.nombre_completo(), font_style2)
                ws.write(row_num, 1, u'%s' % materia.nivel.sesion.nombre, font_style2)
                ws.write(row_num, 2, u'%s' % materia.asignaturamalla.nivelmalla.nombre, font_style2)
                ws.write(row_num, 3, u'%s' % materia.id, font_style2)
                ws.write(row_num, 4, u'%s' % materia.asignatura.nombre, font_style2)
                ws.write(row_num, 5, u'%s' % materia.paralelo, font_style2)
                ws.write(row_num, 6, u'%s' % materiaasignada.matricula.inscripcion.persona.nombre_completo_inverso(), font_style2)
                ws.write(row_num, 7, u'%s' % materiaasignada.matricula.inscripcion.persona.identificacion(), font_style2)
                ws.write(row_num, 8, u'%s' % enrolado, font_style2)
                ws.write(row_num, 9, u'%s' % tienedeuda, font_style2)
                ws.write(row_num, 10, u'%s' % materiaasignada.matricula.get_estado_matricula_display(), font_style2)
                print(u"%s" % materiaasignada)
                row_num += 1
    wb.save(filename)
    print("FIN: ", filename)
# periodo=Periodo.objects.get(id=202)
# cursor = connections['moodle_db'].cursor()
# verificar_enrolado_moodle(periodo,cursor)


# try:
#     periodo=Periodo.objects.get(id=126)
#     listadoprofesormateria = ProfesorMateria.objects.filter(materia__nivel__periodo=periodo, tipoprofesor__id__in=[1, 11, 12, 14], status=True).order_by('materia__asignaturamalla__malla__carrera')
#     for profemateria in listadoprofesormateria:
#         print(u"Procesando %s --- (%s)" % (profemateria, profemateria.id))
#         profemateria.afinidadcampoamplio = False
#         profemateria.afinidadcampoespecifico = False
#         profemateria.afinidadcampodetallado = False
#         if profemateria.materia.asignaturamalla.afinidaddoctorado:
#             if profemateria.profesor.persona.titulacion_set.filter(titulo__grado_id=1, titulo__status=True, status=True):
#                 itemtitulo = profemateria.profesor.persona.titulacion_set.filter(titulo__grado_id=1, titulo__status=True, status=True)[0]
#                 profemateria.tituloafin = itemtitulo.titulo
#                 profemateria.afinidadcampoamplio = True
#                 profemateria.afinidadcampodetallado = True
#                 profemateria.afinidadcampoespecifico = True
#         else:
#             titulos = profemateria.profesor.persona.titulacion_set.filter(titulo__grado_id__in=[1,2,5], titulo__status=True, status=True)
#             if titulos:
#                 eCampos = CamposTitulosPostulacion.objects.filter(status=True, titulo_id__in=titulos.values_list("titulo_id", flat=True))
#                 for eCampo in eCampos:
#                     ecampoamplios=[]
#                     ecampoespecificos=[]
#                     ecampodetallados=[]
#                     if profemateria.materia.asignaturamalla.areaconocimientotitulacion:
#                         ecampoamplios = eCampo.campoamplio.filter(pk=profemateria.materia.asignaturamalla.areaconocimientotitulacion.pk)
#                     if profemateria.materia.asignaturamalla.subareaconocimiento:
#                         ecampoespecificos = eCampo.campoespecifico.filter(pk=profemateria.materia.asignaturamalla.subareaconocimiento.pk)
#                     if profemateria.materia.asignaturamalla.subareaespecificaconocimiento:
#                         ecampodetallados = eCampo.campodetallado.filter(pk=profemateria.materia.asignaturamalla.subareaespecificaconocimiento.pk)
#                     if len(ecampoamplios) > 0:
#                         profemateria.afinidadcampoamplio = True
#                     if len(ecampoespecificos) > 0:
#                         profemateria.afinidadcampoespecifico = True
#                     if len(ecampodetallados) > 0:
#                         profemateria.afinidadcampodetallado = True
#                 for item in titulos.order_by('titulo__grado_id'):
#                     eCampos = CamposTitulosPostulacion.objects.filter(status=True, titulo=item.titulo).order_by('titulo__grado_id')
#                     if eCampos.exists():
#                         eCampo = eCampos.first()
#                         if item.titulo.grado_id in [1,2,5]:
#                             ecampoamplios=[]
#                             ecampoespecificos=[]
#                             ecampodetallados=[]
#                             if profemateria.materia.asignaturamalla.areaconocimientotitulacion:
#                                 ecampoamplios = eCampo.campoamplio.filter(pk=profemateria.materia.asignaturamalla.areaconocimientotitulacion.pk)
#                             if profemateria.materia.asignaturamalla.subareaconocimiento:
#                                 ecampoespecificos = eCampo.campoespecifico.filter(pk=profemateria.materia.asignaturamalla.subareaconocimiento.pk)
#                             if profemateria.materia.asignaturamalla.subareaespecificaconocimiento:
#                                 ecampodetallados = eCampo.campodetallado.filter(pk=profemateria.materia.asignaturamalla.subareaespecificaconocimiento.pk)
#                             if len(ecampoamplios) > 0 or len(ecampoespecificos) > 0 or len(ecampodetallados) > 0:
#                                 profemateria.tituloafin=item.titulo
#                                 break
#         profemateria.save()
#
# except Exception as ex:
#     print('error: %s' % ex)
#
# periodo=Periodo.objects.get(id=119)
# matriculas=Matricula.objects.filter(status=True, nivel__periodo=periodo)
# for matricula in matriculas:
#     modulo_ingles=None
#     guardo_nota=False
#     modulo_ingles = ModuloMalla.objects.values_list('asignatura_id',flat=True).filter(malla=matricula.inscripcion.mi_malla(), status=True).order_by('orden').exclude(asignatura_id=782)
#     for materiaasignada in MateriaAsignada.objects.filter(status=True,retiramateria=False,matricula=matricula, materia__asignatura__in=modulo_ingles,estado_id=3):
#         guardo_nota = False
#         for notasmooc in materiaasignada.materia.notas_de_moodle(materiaasignada.matricula.inscripcion.persona):
#             campo = materiaasignada.campo(notasmooc[1].upper())
#             if Decimal(notasmooc[0]) >= 70 and materiaasignada.notafinal <= 0:
#                 if type(notasmooc[0]) is Decimal:
#                     if null_to_decimal(campo.valor) != float(notasmooc[0]):
#                         actualizar_nota_planificacion(materiaasignada.id, notasmooc[1].upper(), notasmooc[0])
#                         auditorianotas = AuditoriaNotas(evaluaciongenerica=campo, manual=False,
#                                                         calificacion=notasmooc[0])
#                         auditorianotas.save()
#                 else:
#                     if null_to_decimal(campo.valor) != float(0):
#                         actualizar_nota_planificacion(materiaasignada.id, notasmooc[1].upper(), notasmooc[0])
#                         auditorianotas = AuditoriaNotas(evaluaciongenerica=campo, manual=False, calificacion=0)
#                         auditorianotas.save()
#                 guardo_nota=True
#                 print(u"Importa nota %s"%materiaasignada)
#         if guardo_nota:
#             materiaasignada.importa_nota = True
#             materiaasignada.cerrado = True
#             materiaasignada.fechacierre = datetime.now().date()
#             materiaasignada.save()
#             d = locals()
#             exec(materiaasignada.materia.modeloevaluativo.logicamodelo, globals(), d)
#             d['calculo_modelo_evaluativo'](materiaasignada)
#             materiaasignada.cierre_materia_asignada()



# import xlwt
# from xlwt import *
# from django.http import HttpResponse
#
# response = HttpResponse(content_type="application/ms-excel")
# response['Content-Disposition'] = 'attachment; filename=reporte_ingles.xls'
# style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
# style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
# style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
# title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
# style1 = easyxf(num_format_str='D-MMM-YY')
# font_style = XFStyle()
# font_style.font.bold = True
# font_style2 = XFStyle()
# font_style2.font.bold = False
# wb = xlwt.Workbook()
# ws = wb.add_sheet('Sheetname')
# estilo = xlwt.easyxf('font: height 350, name Arial, colour_index black, bold on, italic on; align: wrap on, vert centre, horiz center;')
# ws.write_merge(0, 0, 0, 9, 'UNIVERSIDAD ESTATAL DE MILAGRO', estilo)
# output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
# nombre = "Lista" + datetime.now().strftime('%Y%m%d_%H%M%S') + ".xls"
# filename = os.path.join(output_folder, nombre)
# columns = [(u"materia asignada", 6000),
#            (u"estado", 6000),
#            (u"nota sga", 6000),
#            (u"nota moodle", 6000),
#            ]
# row_num = 3
# for col_num in range(len(columns)):
#     ws.write(row_num, col_num, columns[col_num][0], font_style)
#     ws.col(col_num).width = columns[col_num][1]
# row_num = 4
# periodo=Periodo.objects.get(id=119)
# matriculas=Matricula.objects.filter(status=True, nivel__periodo=periodo)
# for matricula in matriculas:
#     modulo_ingles=None
#     guardo_nota=False
#     modulo_ingles = ModuloMalla.objects.values_list('asignatura_id',flat=True).filter(malla=matricula.inscripcion.mi_malla(), status=True).order_by('orden').exclude(asignatura_id=782)
#     for materiaasignada in MateriaAsignada.objects.filter(matricula=matricula, materia__asignatura__in=modulo_ingles,estado_id=3):
#         guardo_nota = False
#         ws.write(row_num, 0, u'%s' % materiaasignada, font_style2)
#         ws.write(row_num, 1, u'%s' % materiaasignada.estado, font_style2)
#         ws.write(row_num, 2, u'%s' % materiaasignada.notafinal, font_style2)
#         for notasmooc in materiaasignada.materia.notas_de_moodle(materiaasignada.matricula.inscripcion.persona):
#             ws.write(row_num, 3, u'%s' % notasmooc[0], font_style2)
#         print(u"%s" % materiaasignada)
#         row_num += 1
# wb.save(filename)
# print("FIN: ", filename)


# RETIRAR DE MODULOS

# periodo=Periodo.objects.get(id=119)
# matriculas=Matricula.objects.filter(status=True, nivel__periodo=periodo,id=396564)
# cont=1
# for matricula in matriculas:
#     modulo_ingles=None
#     modulo_ingles = ModuloMalla.objects.values_list('asignatura_id',flat=True).filter(malla=matricula.inscripcion.mi_malla(), status=True).order_by('orden').exclude(asignatura_id=782)
#     for materiaasignada in MateriaAsignada.objects.filter(status=True,retiramateria=False,matricula=matricula, materia__asignatura__in=modulo_ingles,estado_id=3):
#         if materiaasignada.notafinal <= 0:
#             if not MateriaAsignadaRetiro.objects.filter(status=True).exists():
#                 retiro = MateriaAsignadaRetiro(materiaasignada=materiaasignada,
#                                                motivo='RETIRO POR TÉRMINO DEl PROCESO DE INGLÉS',
#                                                valida=False,
#                                                fecha=datetime.now().date())
#                 retiro.save()
#             if not materiaasignada.retiramateria:
#                 materiaasignada.retiramateria = True
#                 materiaasignada.save()
#             print(u"%s --- %s" % (materiaasignada,cont))
#             cont+=1


# PROCESO DE MATRICULA INGLÉS ESTUDIANTE

# import xlwt
# from xlwt import *
# from django.http import HttpResponse
#
# response = HttpResponse(content_type="application/ms-excel")
# response['Content-Disposition'] = 'attachment; filename=reporte_matricula_modulo_ingles.xls'
# style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
# style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
# style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
# title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
# style1 = easyxf(num_format_str='D-MMM-YY')
# font_style = XFStyle()
# font_style.font.bold = True
# font_style2 = XFStyle()
# font_style2.font.bold = False
# wb = xlwt.Workbook()
# ws = wb.add_sheet('Sheetname')
# estilo = xlwt.easyxf('font: height 350, name Arial, colour_index black, bold on, italic on; align: wrap on, vert centre, horiz center;')
# ws.write_merge(0, 0, 0, 9, 'UNIVERSIDAD ESTATAL DE MILAGRO', estilo)
# output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
# nombre = "Lista" + datetime.now().strftime('%Y%m%d_%H%M%S') + ".xls"
# filename = os.path.join(output_folder, nombre)
# columns = [(u"Id", 6000),
#            (u"carrera", 6000),
#            (u"cedula", 6000),
#            (u"alumno", 6000),
#            (u"modulo", 6000),
#            (u"cant matricula", 6000),
#            (u"materia asignada", 6000),
#            (u"perdida de gratuidad", 6000),
#            ]
# row_num = 3
# for col_num in range(len(columns)):
#     ws.write(row_num, col_num, columns[col_num][0], font_style)
#     ws.col(col_num).width = columns[col_num][1]
# row_num = 4
#
# contador=1
# persona_responsable=Persona.objects.get(id=20539)
# for matricula in Matricula.objects.filter(status=True, nivel__periodo_id=126,retiradomatricula=False, id__in=[435812,
#                                                                                                             436184,
#                                                                                                             445051,
#                                                                                                             440503,
#                                                                                                             440413,
#                                                                                                             439828,
#                                                                                                             452219,
#                                                                                                             474735,
#                                                                                                             449442,
#                                                                                                             436158,
#                                                                                                             441180,
#                                                                                                             453820,
#                                                                                                             441343,
#                                                                                                             451801,
#                                                                                                             439794,
#                                                                                                             440332,
#                                                                                                             440420,
#                                                                                                             446148,
#                                                                                                             474711,
#                                                                                                             451811,
#                                                                                                             442949,
#                                                                                                             440656]).exclude(inscripcion__carrera__id__in=[7,138,129,90,157]):
#     try:
#         bandera=0
#         modulomalla=None
#         ws.write(row_num, 0, u'%s' % matricula.id, font_style2)
#         ws.write(row_num, 1, u'%s' % matricula.inscripcion.carrera, font_style2)
#         ws.write(row_num, 2, u'%s' % matricula.inscripcion.persona.identificacion(), font_style2)
#         ws.write(row_num, 3, u'%s' % matricula.inscripcion.persona, font_style2)
#         id_modulos_ingles_todos=None
#         if matricula.inscripcion.modulos_ingles_mi_malla():
#             id_modulos_ingles_todos=matricula.inscripcion.modulos_ingles_mi_malla().values_list('asignatura_id',flat=True).distinct()
#         if id_modulos_ingles_todos:
#             if matricula.inscripcion.ultimo_modulo_ingles_pendiente():
#                 modulomalla=matricula.inscripcion.ultimo_modulo_ingles_pendiente()
#                 materias=Materia.objects.filter(status=True, nivel_id=719, asignatura=modulomalla.asignatura,inglesepunemi=True )
#                 for materia in materias:
#                     if materia.tiene_capacidad():
#                         matriculas=1
#                         materiaasignada=None
#                         matriculas = matricula.inscripcion.historicorecordacademico_set.filter(asignatura=materia.asignatura).count() + 1
#                         if not MateriaAsignada.objects.filter(matricula=matricula, materia__asignatura=modulomalla.asignatura).exists():
#                             materiaasignada = MateriaAsignada(matricula=matricula,
#                                                               materia=materia,
#                                                               notafinal=0,
#                                                               asistenciafinal=0,
#                                                               cerrado=False,
#                                                               matriculas=matriculas,
#                                                               observaciones='',
#                                                               estado_id=NOTA_ESTADO_EN_CURSO,
#                                                               automatricula=True,
#                                                               importa_nota=False)
#                             materiaasignada.save()
#                             materiaasignada.evaluacion()
#                             creditos = materiaasignada.materia.creditos
#                             if materiaasignada.existe_modulo_en_malla():
#                                 creditos = materiaasignada.materia_modulo_malla().creditos
#                             registro = AgregacionEliminacionMaterias(matricula=matricula,
#                                                                      agregacion=True,
#                                                                      asignatura=materiaasignada.materia.asignatura,
#                                                                      responsable=persona_responsable,
#                                                                      fecha=datetime.now().date(),
#                                                                      creditos=creditos,
#                                                                      nivelmalla=materiaasignada.materia.nivel.nivelmalla if materiaasignada.materia.nivel.nivelmalla else None,
#                                                                      matriculas=materiaasignada.matriculas)
#                             registro.save()
#                             if str(matricula.inscripcion_id) not in variable_valor('INSCRIPCIONES_INGLES'):
#                                 if matriculas>1 or matricula.inscripcion.persona.tiene_otro_titulo(matricula.inscripcion):
#                                     matricula.nuevo_calculo_matricula_ingles(materiaasignada)
#                             ws.write(row_num, 4, u'%s' % modulomalla, font_style2)
#                             ws.write(row_num, 5, u'%s' % matriculas, font_style2)
#                             ws.write(row_num, 6, u'%s' % materiaasignada, font_style2)
#                             ws.write(row_num, 7, u'%s' % "SI" if matricula.inscripcion.persona.tiene_otro_titulo(matricula.inscripcion) else "NO", font_style2)
#                             row_num += 1
#                             print(u"(%s) -- %s [%s]" % (contador, materiaasignada, materiaasignada.matricula_id))
#                             contador += 1
#                             bandera=1
#                             break
#                         else:
#                             ws.write(row_num, 4, u'YA TIENE MATRICULA EN EL MODULO %s' %materia, font_style2)
#                             ws.write(row_num, 5, u'YA TIENE MATRICULA EN EL MODULO %s' %materia, font_style2)
#                             ws.write(row_num, 6, u'YA TIENE MATRICULA EN EL MODULO %s' %materia, font_style2)
#                             ws.write(row_num, 7, u'YA TIENE MATRICULA EN EL MODULO %s' %materia, font_style2)
#                             row_num += 1
#                             print(u"(%s) -- YA TIENE MATRICULA EN EL MODULO %s" % (contador,materia))
#                             contador += 1
#                             bandera = 1
#             else:
#                 ws.write(row_num, 4, u'NO AY ULTIMO MÓDULO DE INGLÉS PENDIENTE', font_style2)
#                 ws.write(row_num, 5, u'NO AY ULTIMO MÓDULO DE INGLÉS PENDIENTE', font_style2)
#                 ws.write(row_num, 6, u'NO AY ULTIMO MÓDULO DE INGLÉS PENDIENTE', font_style2)
#                 ws.write(row_num, 7, u'NO AY ULTIMO MÓDULO DE INGLÉS PENDIENTE', font_style2)
#                 row_num += 1
#                 print(u"(%s) -- NO AY ULTIMO MÓDULO DE INGLÉS PENDIENTE %s" % (contador,matricula))
#                 contador += 1
#                 bandera = 1
#             # else:
#             #     ws.write(row_num, 4, u'YA TIENE MATRICULA EN AL MENOS UN MODULO %s' %matricula, font_style2)
#             #     ws.write(row_num, 5, u'YA TIENE MATRICULA EN AL MENOS UN MODULO %s' %matricula, font_style2)
#             #     ws.write(row_num, 6, u'YA TIENE MATRICULA EN AL MENOS UN MODULO %s' %matricula, font_style2)
#             #     ws.write(row_num, 7, u'YA TIENE MATRICULA EN AL MENOS UN MODULO %s' %matricula, font_style2)
#             #     row_num += 1
#             #     print(u"(%s) -- YA TIENE MATRICULA EN AL MENOS UN MODULO %s" %(matricula, contador))
#             #     contador += 1
#             #     bandera = 1
#         else:
#             ws.write(row_num, 4, u'NO HAY MODULOS DE INGLES EN LA MALLA', font_style2)
#             ws.write(row_num, 5, u'NO HAY MODULOS DE INGLES EN LA MALLA', font_style2)
#             ws.write(row_num, 6, u'NO HAY MODULOS DE INGLES EN LA MALLA', font_style2)
#             ws.write(row_num, 7, u'NO HAY MODULOS DE INGLES EN LA MALLA', font_style2)
#             row_num += 1
#             print(u"(%s) -- NO HAY MODULOS DE INGLES EN LA MALLA" % contador)
#             contador += 1
#             bandera = 1
#         if bandera == 0:
#             ws.write(row_num, 4, u'NO SE MATRICULA - NO HAY CUPO %s -- %s '%(modulomalla,matricula), font_style2)
#             ws.write(row_num, 5, u'NO SE MATRICULA - NO HAY CUPO %s -- %s '%(modulomalla,matricula), font_style2)
#             ws.write(row_num, 6, u'NO SE MATRICULA - NO HAY CUPO %s -- %s '%(modulomalla,matricula), font_style2)
#             ws.write(row_num, 7, u'NO SE MATRICULA - NO HAY CUPO %s -- %s '%(modulomalla,matricula), font_style2)
#             row_num += 1
#             print(u"NO SE MATRICULA - NO HAY CUPO %s -- %s" % (modulomalla,matricula))
#             contador += 1
#     except Exception as ex:
#         ws.write(row_num, 4, u'SALE ERROR %s' % ( matricula), font_style2)
#         ws.write(row_num, 5, u'SALE ERROR %s' % ( matricula), font_style2)
#         ws.write(row_num, 6, u'SALE ERROR %s' % ( matricula), font_style2)
#         ws.write(row_num, 7, u'SALE ERROR %s' % ( matricula), font_style2)
#         row_num += 1
#         contador += 1
#         print('error: %s %s' % (ex,matricula))
#         pass
# wb.save(filename)
# print("FIN: ", filename)

# PROCESO DE CIERRE DE MATRICULA DE INGLES
def proceso_cierre_ingles(periodo):
    from django.http import HttpResponse
    response = HttpResponse(content_type="application/ms-excel")
    response['Content-Disposition'] = 'attachment; filename=reporte_notas_ingles.xls'
    style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
    style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
    style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
    title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
    style1 = easyxf(num_format_str='D-MMM-YY')
    font_style = XFStyle()
    font_style.font.bold = True
    font_style2 = XFStyle()
    font_style2.font.bold = False
    wb = xlwt.Workbook()
    ws = wb.add_sheet('Sheetname')
    estilo = xlwt.easyxf('font: height 350, name Arial, colour_index black, bold on, italic on; align: wrap on, vert centre, horiz center;')
    ws.write_merge(0, 0, 0, 9, 'UNIVERSIDAD ESTATAL DE MILAGRO', estilo)
    output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
    nombre = "Lista" + datetime.now().strftime('%Y%m%d_%H%M%S') + ".xls"
    filename = os.path.join(output_folder, nombre)
    columns = [(u"CEDULA", 6000),
               (u"APELLIDOS Y NOMBRES", 6000),
               (u"CARRERA", 6000),
               (u"MODULO", 6000),
               (u"NOTA BUCKINGHAM", 6000),
               (u"NOTA MATERIA", 6000),
               (u"ESTADO", 6000),
               (u"CURSO", 6000)
               ]
    row_num = 3
    for col_num in range(len(columns)):
        ws.write(row_num, col_num, columns[col_num][0], font_style)
        ws.col(col_num).width = columns[col_num][1]
    row_num = 4
    cont=0
    materiasasignadas=MateriaAsignada.objects.filter(status=True,materia__inglesepunemi=True, retiramateria=True,materia__nivel__periodo=periodo,materia__nivel_id=781,materia__status=True).order_by('matricula__inscripcion__carrera')
    print(u"%s"%(materiasasignadas.values('id').count()))
    for materiaasignada in materiasasignadas:
        with transaction.atomic():
            try:
                idcursomoodle = materiaasignada.materia.idcursomoodle
                url = 'https://upei.buckcenter.edu.ec/usernamecoursetograde.php?username=%s&curso=%s' % (materiaasignada.matricula.inscripcion.persona.identificacion(),idcursomoodle)
                cont += 1
                ws.write(row_num, 0, materiaasignada.matricula.inscripcion.persona.identificacion())
                ws.write(row_num, 1, materiaasignada.matricula.inscripcion.persona.apellido1 + ' ' + materiaasignada.matricula.inscripcion.persona.apellido2 + ' ' + materiaasignada.matricula.inscripcion.persona.nombres)
                ws.write(row_num, 2, str(materiaasignada.matricula.inscripcion.carrera))
                req = Request(url)
                response = urlopen(req)
                result = json.loads(response.read().decode())
                idcurso = int(result['idcurso'])
                print(u"----- %s -----" % cont)
                print(u"PROCESANDO - %s" % materiaasignada)
                print(u"%s" % result)
                print(u"ID CURSO: %s" % idcurso)
                if idcurso != idcursomoodle:
                    ws.write(row_num, 3, u"%s" % materiaasignada.materia)
                    ws.write(row_num, 4, u"NO COINCIDE CURSO")
                    ws.write(row_num, 5, u"NO COINCIDE CURSO")
                    ws.write(row_num, 6, u"NO COINCIDE CURSO")
                    ws.write(row_num, 7, u"NO COINCIDE CURSO")
                else:
                    try:
                        nota = null_to_decimal(result['nota'], 0)
                        # if nota != materiaasignada.notafinal:
                        #     campo = materiaasignada.campo('EX')
                        #     actualizar_nota_planificacion(materiaasignada.id, 'EX', nota)
                        #     auditorianotas = AuditoriaNotas(evaluaciongenerica=campo, manual=False, calificacion=nota)
                        #     auditorianotas.save()
                        #     materiaasignada.importa_nota = True
                        #     materiaasignada.cerrado = True
                        #     materiaasignada.fechacierre = datetime.now().date()
                        #     materiaasignada.save()
                        #     d = locals()
                        #     exec(materiaasignada.materia.modeloevaluativo.logicamodelo, globals(), d)
                        #     d['calculo_modelo_evaluativo'](materiaasignada)
                        #     materiaasignada.cierre_materia_asignada()
                        #     print(u"IMPORTA Y CIERRA -- %s" % (materiaasignada))
                        # ws.write(row_num, 3, u"%s" % materiaasignada.materia)
                        # ws.write(row_num, 4, u"%s" % result['nota'])
                        # ws.write(row_num, 5, nota)
                        # ws.write(row_num, 6, u"APROBADO" if nota >= 70 else "REPROBADO")
                        # ws.write(row_num, 7, u"COINCIDE CURSO" if idcurso == materiaasignada.materia.idcursomoodle else "NO COINCIDE CURSO")
                    except:
                        nota=result['nota']
                        if result['nota'] == '-':
                            # if not MateriaAsignadaRetiro.objects.filter(status=True, materiaasignada=materiaasignada).exists():
                            #     retiro = MateriaAsignadaRetiro(materiaasignada=materiaasignada,
                            #                                    motivo='RETIRO POR TÉRMINO DEl PROCESO DE INGLÉS 1S 2023 BUCKINGHAM',
                            #                                    valida=False,
                            #                                    fecha=datetime.now().date())
                            #     retiro.save()
                            # if not materiaasignada.retiramateria:
                            #     materiaasignada.retiramateria = True
                            #     materiaasignada.save()
                            rubros=materiaasignada.rubro.filter(status=True,observacion='INGLÉS NOVIEMBRE 2022 MARZO 2023')
                            for rubro in rubros:
                                if not rubro.pagos():
                                    # rubro.delete()
                                    print(u"ELIMINADO -- %s" % (materiaasignada))
                        ws.write(row_num, 3, u"%s" % materiaasignada.materia)
                        ws.write(row_num, 4, u"%s" % result['nota'])
                        ws.write(row_num, 5, nota)
                        ws.write(row_num, 6, u"RETIRADO")
                        ws.write(row_num, 7,u"COINCIDE CURSO" if idcurso == materiaasignada.materia.idcursomoodle else "NO COINCIDE CURSO")
            except Exception as ex:
                transaction.set_rollback(True)
                print('error: %s' % (ex))
                ws.write(row_num, 3, u"%s" % ex)
                ws.write(row_num, 4, u"%s" % ex)
                ws.write(row_num, 5, u"%s" % ex)
                ws.write(row_num, 6, u"%s" % ex)
                ws.write(row_num, 7, u"%s" % ex)
                pass
            row_num += 1
    wb.save(filename)
    print("FIN: ", filename)

# periodo=Periodo.objects.get(id=153)
# proceso_cierre_ingles(periodo)
#
# try:
#     miarchivo = openpyxl.load_workbook("FORMATO PARA CONFIGURACIÓN DE MIGRACION DE MALLA 11-10-2022 REVISADA.xlsx")
#     lista = miarchivo.get_sheet_by_name('MALLA 2022')
#     totallista = lista.rows
#     a=0
#     for filas in totallista:
#         a += 1
#         if a > 2:
#             idasignaturamalla2018 = int(filas[6].value)
#             idasignaturamalla2022=int(filas[0].value)
#             if not TablaEquivalenciaAsignaturas.objects.filter(status=True,asignaturamalla_id=idasignaturamalla2018).exists():
#                 tablaeq=TablaEquivalenciaAsignaturas(asignaturamalla_id=idasignaturamalla2018,
#                                              asignaturamallasalto_id=idasignaturamalla2022)
#                 tablaeq.save()
#                 print(u"INSERTA EQUIVALENCIA %s"%tablaeq)
#             else:
#                 tablaeq=TablaEquivalenciaAsignaturas.objects.filter(status=True, asignaturamalla_id=idasignaturamalla2018)[0]
#                 tablaeq.asignaturamallasalto_id=idasignaturamalla2022
#                 tablaeq.save()
#                 print(u"ACTUALIZA EQUIVALENCIA %s" % tablaeq)
#             print(u"Fila %s"%a)
# except Exception as ex:
#         print('error: %s' % ex)



# MIGRAR ASIGNATURAS DE RECORD A NUEVA MALLA ENFERMERIA
# import xlwt
# from xlwt import *
# from django.http import HttpResponse
#
# response = HttpResponse(content_type="application/ms-excel")
# response['Content-Disposition'] = 'attachment; filename=reporte_notas_ingles.xls'
# style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
# style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
# style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
# title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
# style1 = easyxf(num_format_str='D-MMM-YY')
# font_style = XFStyle()
# font_style.font.bold = True
# font_style2 = XFStyle()
# font_style2.font.bold = False
# wb = xlwt.Workbook()
# ws = wb.add_sheet('Sheetname')
# estilo = xlwt.easyxf('font: height 350, name Arial, colour_index black, bold on, italic on; align: wrap on, vert centre, horiz center;')
# ws.write_merge(0, 0, 0, 9, 'UNIVERSIDAD ESTATAL DE MILAGRO', estilo)
# output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
# nombre = "Lista" + datetime.now().strftime('%Y%m%d_%H%M%S') + ".xls"
# filename = os.path.join(output_folder, nombre)
# columns = [(u"CEDULA", 6000),
#            (u"APELLIDOS Y NOMBRES", 6000),
#            (u"OBSERVACIÓN", 6000)
#            ]
# row_num = 3
# for col_num in range(len(columns)):
#     ws.write(row_num, col_num, columns[col_num][0], font_style)
#     ws.col(col_num).width = columns[col_num][1]
# row_num = 4
#
# matriculadosenfermeria=Matricula.objects.filter(status=True,
#                                                 inscripcion__carrera_id=110,
#                                                 nivel__periodo_id=126,id=442938,
#                                                 retiradomatricula=False,
#                                                 inscripcion__inscripcionnivel__nivel__orden__lt=7).order_by('inscripcion__persona__apellido1', 'inscripcion__persona__apellido2', 'inscripcion__persona__nombres')
# print(u"%s"%matriculadosenfermeria.count())
# cont=0
# for matricula in matriculadosenfermeria:
#     try:
#         cont+=1
#         matricula.pasoayuda=True
#         matricula.save()
#         print(u"%s - %s"%(matricula,cont))
#         inscripcion=matricula.inscripcion
#         ws.write(row_num, 0, matricula.inscripcion.persona.identificacion())
#         ws.write(row_num, 1, matricula.inscripcion.persona.nombre_completo())
#         ws.write(row_num, 2, str(matricula.inscripcion.carrera))
#         if InscripcionMalla.objects.filter(status=True,inscripcion=inscripcion,  malla_id=173).exists():
#             imantigua=InscripcionMalla.objects.filter(status=True,inscripcion=inscripcion,  malla_id=173)[0]
#             imantigua.status=False
#             imantigua.save()
#             print(u"Desactiva antigua inscripcion -----------------------------")
#         if not InscripcionMalla.objects.filter(status=True,inscripcion=inscripcion,  malla_id=410).exists():
#             imnueva = InscripcionMalla(inscripcion=inscripcion, malla_id=410)
#             imnueva.save()
#             print(u"Crea nueva inscripcion -----------------------------")
#         equivalencias=TablaEquivalenciaAsignaturas.objects.filter(status=True,id=1442).order_by('asignaturamallasalto__nivelmalla__orden')
#         for equivalencia in equivalencias:
#             print(u"2022 - %s"%equivalencia.asignaturamallasalto)
#             recordantiguo=inscripcion.recordacademico_set.filter(id=2946150)
#             if recordantiguo.exists():
#                 print(u"2018 - %s" % equivalencia.asignaturamalla)
#                 recordantiguo=recordantiguo[0]
#                 print(u"Record antiguo: %s" % recordantiguo)
#                 recordnuevo=None
#                 recordantiguo.status = False
#                 recordantiguo.save(update_asignaturamalla=False)
#                 if not RecordAcademico.objects.filter(status=True,inscripcion=inscripcion, asignaturamalla=equivalencia.asignaturamallasalto).exists():
#                     recordnuevo=RecordAcademico(inscripcion=inscripcion,
#                                                 matriculas=recordantiguo.matriculas,
#                                                 asignaturamalla=equivalencia.asignaturamallasalto,
#                                                 asignatura=equivalencia.asignaturamallasalto.asignatura,
#                                                 asignaturaold_id=3209,
#                                                 nota=recordantiguo.nota,
#                                                 asistencia=recordantiguo.asistencia,
#                                                 sinasistencia=recordantiguo.sinasistencia,
#                                                 fecha=recordantiguo.fecha,
#                                                 noaplica=recordantiguo.noaplica,
#                                                 aprobada=recordantiguo.aprobada,
#                                                 convalidacion=recordantiguo.convalidacion,
#                                                 pendiente=recordantiguo.pendiente,
#                                                 creditos=equivalencia.asignaturamallasalto.creditos,
#                                                 horas=equivalencia.asignaturamallasalto.horas,
#                                                 valida=recordantiguo.valida,
#                                                 validapromedio=recordantiguo.validapromedio,
#                                                 observaciones=recordantiguo.observaciones+" con base a RESOLUCIÓN OCAS-SO-7-2022-No16",
#                                                 materiaregular=recordantiguo.materiaregular,
#                                                 materiacurso=None,
#                                                 completonota=recordantiguo.completonota,
#                                                 completoasistencia=recordantiguo.completoasistencia,
#                                                 fechainicio=recordantiguo.fechainicio,
#                                                 fechafin=recordantiguo.fechafin,
#                                                 suficiencia=recordantiguo.suficiencia,
#                                                 asignaturamallahistorico_id=4875,
#                                                 reverso=False)
#                     recordnuevo.save()
#                     print(u"Crea nuevo record %s"%recordnuevo)
#                 elif RecordAcademico.objects.filter(status=True, inscripcion=inscripcion, asignaturamalla=equivalencia.asignaturamallasalto):
#                     recordnuevo=RecordAcademico.objects.filter(status=True, inscripcion=inscripcion, asignaturamalla=equivalencia.asignaturamallasalto)[0]
#                     recordnuevo.matriculas = recordantiguo.matriculas
#                     recordnuevo.asignaturamalla = equivalencia.asignaturamallasalto
#                     recordnuevo.asignatura = equivalencia.asignaturamallasalto.asignatura
#                     recordnuevo.asignaturaold = recordantiguo.asignaturamalla.asignatura
#                     recordnuevo.nota = recordantiguo.nota
#                     recordnuevo.asistencia = recordantiguo.asistencia
#                     recordnuevo.sinasistencia = recordantiguo.sinasistencia
#                     recordnuevo.fecha = recordantiguo.fecha
#                     recordnuevo.noaplica = recordantiguo.noaplica
#                     recordnuevo.aprobada = recordantiguo.aprobada
#                     recordnuevo.convalidacion = recordantiguo.convalidacion
#                     recordnuevo.pendiente = recordantiguo.pendiente
#                     recordnuevo.creditos = equivalencia.asignaturamallasalto.creditos
#                     recordnuevo.horas = equivalencia.asignaturamallasalto.horas
#                     recordnuevo.valida = recordantiguo.valida
#                     recordnuevo.validapromedio = recordantiguo.validapromedio
#                     recordnuevo.observaciones = recordantiguo.observaciones + " con base a RESOLUCIÓN OCAS-SO-7-2022-No16"
#                     recordnuevo.materiaregular = recordantiguo.materiaregular
#                     recordnuevo.materiacurso = None
#                     recordnuevo.completonota = recordantiguo.completonota
#                     recordnuevo.completoasistencia = recordantiguo.completoasistencia
#                     recordnuevo.fechainicio = recordantiguo.fechainicio
#                     recordnuevo.fechafin = recordantiguo.fechafin
#                     recordnuevo.suficiencia = recordantiguo.suficiencia
#                     recordnuevo.asignaturamallahistorico = recordantiguo.asignaturamalla
#                     recordnuevo.reverso = False
#                     recordnuevo.save()
#                 if recordnuevo:
#                     historicos=HistoricoRecordAcademico.objects.filter(status=True,recordacademico=recordantiguo).update(recordacademico=recordnuevo)
#                     respaldo = RespaldoRecordAcademico.objects.filter(status=True, recordacademicooriginal=recordantiguo)
#                     if not respaldo.exists():
#                         respaldorecord = RespaldoRecordAcademico(
#                             recordacademicooriginal=recordantiguo,
#                             recordacademiconuevo=recordnuevo
#                         )
#                         respaldorecord.save()
#                     else:
#                         respaldorecord = respaldo[0]
#                         respaldorecord.recordacademiconuevo = recordnuevo
#                         respaldorecord.save()
#                     print(u"Record actualizado %s" % recordnuevo)
#             else:
#                 ws.write(row_num, 3, "NO ENCONTRO RECORD ANTIGUO %s" %equivalencia.asignaturamalla)
#                 row_num += 1
#         time.sleep(3)
#     except Exception as ex:
#         print('error: %s' % ex)
#         ws.write(row_num, 3, str(ex))
#         row_num += 1
#
# wb.save(filename)
# print("FIN: ", filename)



# matriculadosenfermeria=Matricula.objects.filter(status=True,
#                                                 inscripcion__carrera_id=110,
#                                                 nivel__periodo_id=126,
#                                                 retiradomatricula=False,
#                                                 inscripcion__inscripcionnivel__nivel__orden__lt=7).order_by('inscripcion__persona__apellido1', 'inscripcion__persona__apellido2', 'inscripcion__persona__nombres')
# print(u"%s"%matriculadosenfermeria.count())
# cont=0
# for matricula in matriculadosenfermeria:
#     try:
#         cont+=1
#         print(u"%s - %s"%(matricula,cont))
#         inscripcion=matricula.inscripcion
#         records=inscripcion.recordacademico_set.filter(status=True,asignaturamalla__isnull=False, observaciones__icontains='CON BASE A RESOLUCIÓN OCAS-SO-7-2022-NO16')
#         for record in records:
#             print(u"2022 - %s"%record)
#             if record.materiaregular.asignaturamalla!=record.asignaturamalla:
#                 emateria=record.materiaregular
#                 emateria.asignatura= record.asignatura
#                 emateria.asignaturaold= record.asignaturamallahistorico.asignatura
#                 emateria.asignaturamalla= record.asignaturamalla
#                 emateria.save()
#             respaldo = RespaldoRecordAcademico.objects.filter(status=True, recordacademiconuevo=record)
#             if respaldo.exists():
#                 respaldorecord = respaldo[0]
#                 historicos = HistoricoRecordAcademico.objects.filter(Q(recordacademico=respaldorecord.recordacademicooriginal) | Q(recordacademico=record), status=True)
#                 for x in historicos:
#                     x.recordacademico = record
#                     x.asignaturamalla = record.asignaturamalla
#                     x.asignatura = record.asignatura
#                     x.asignaturaold = record.asignaturaold
#                     x.save(update_asignaturamalla=False)
#     except Exception as ex:
#         print('error: %s' % ex)


#
# eMatriculas = Matricula.objects.filter(status=True, id__in=[558854,558863])
# total = len(eMatriculas.values("id"))
# print(f"Total encontrados: {total}")
# contador = 0
# for eMatricula in eMatriculas:
#     contador += 1
#     print(f"Inicia --> ({total}/{contador}) - {eMatricula.__str__()}")
#     # eMatricula.calcula_nivel()
#     eMatricula.agregacion_aux(None)
#     print(f"Finalizo --> ({total}/{contador}) - {eMatricula.__str__()}")
# print("---------------------------------------------------------------------------------------------------------")




# PROCESO DE MATRICULA INGLÉS ESTUDIANTE

# import xlwt
# from xlwt import *
# from django.http import HttpResponse
#
# response = HttpResponse(content_type="application/ms-excel")
# response['Content-Disposition'] = 'attachment; filename=reporte_matricula_modulo_ingles.xls'
# style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
# style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
# style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
# title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
# style1 = easyxf(num_format_str='D-MMM-YY')
# font_style = XFStyle()
# font_style.font.bold = True
# font_style2 = XFStyle()
# font_style2.font.bold = False
# wb = xlwt.Workbook()
# ws = wb.add_sheet('Sheetname')
# estilo = xlwt.easyxf('font: height 350, name Arial, colour_index black, bold on, italic on; align: wrap on, vert centre, horiz center;')
# ws.write_merge(0, 0, 0, 9, 'UNIVERSIDAD ESTATAL DE MILAGRO', estilo)
# output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
# nombre = "Lista" + datetime.now().strftime('%Y%m%d_%H%M%S') + ".xls"
# filename = os.path.join(output_folder, nombre)
# columns = [(u"Id", 6000),
#            (u"carrera", 6000),
#            (u"cedula", 6000),
#            (u"alumno", 6000),
#            (u"modulo", 6000),
#            (u"cant matricula", 6000),
#            (u"materia asignada", 6000),
#            (u"perdida de gratuidad", 6000),
#            ]
# row_num = 3
# for col_num in range(len(columns)):
#     ws.write(row_num, col_num, columns[col_num][0], font_style)
#     ws.col(col_num).width = columns[col_num][1]
# row_num = 4
#
# contador=1
# persona_responsable=Persona.objects.get(id=20539)
# for matricula in Matricula.objects.filter(status=True, nivel__periodo_id=153,retiradomatricula=False).exclude(inscripcion__carrera__id__in=[7,138,129,90,157]):
#     try:
#         bandera=0
#         modulomalla=None
#         ws.write(row_num, 0, u'%s' % matricula.id, font_style2)
#         ws.write(row_num, 1, u'%s' % matricula.inscripcion.carrera, font_style2)
#         ws.write(row_num, 2, u'%s' % matricula.inscripcion.persona.identificacion(), font_style2)
#         ws.write(row_num, 3, u'%s' % matricula.inscripcion.persona, font_style2)
#         id_modulos_ingles_todos=None
#         if matricula.inscripcion.modulos_ingles_mi_malla():
#             id_modulos_ingles_todos=matricula.inscripcion.modulos_ingles_mi_malla().values_list('asignatura_id',flat=True).distinct()
#         if id_modulos_ingles_todos:
#             if matricula.inscripcion.ultimo_modulo_ingles_pendiente():
#                 modulomalla=matricula.inscripcion.ultimo_modulo_ingles_pendiente()
#                 materias=Materia.objects.filter(status=True, nivel_id=764, asignatura=modulomalla.asignatura,inglesepunemi=True ).exclude(id__in=[60950,60949])
#                 for materia in materias:
#                     if materia.tiene_capacidad():
#                         matriculas=1
#                         materiaasignada=None
#                         matriculas = matricula.inscripcion.historicorecordacademico_set.filter(asignatura=materia.asignatura).count() + 1
#                         if not MateriaAsignada.objects.filter(matricula=matricula, materia__asignatura=modulomalla.asignatura).exists():
#                             materiaasignada = MateriaAsignada(matricula=matricula,
#                                                               materia=materia,
#                                                               notafinal=0,
#                                                               asistenciafinal=0,
#                                                               cerrado=False,
#                                                               matriculas=matriculas,
#                                                               observaciones='',
#                                                               estado_id=NOTA_ESTADO_EN_CURSO,
#                                                               automatricula=True,
#                                                               importa_nota=False)
#                             materiaasignada.save()
#                             materiaasignada.evaluacion()
#                             creditos = materiaasignada.materia.creditos
#                             if materiaasignada.existe_modulo_en_malla():
#                                 creditos = materiaasignada.materia_modulo_malla().creditos
#                             registro = AgregacionEliminacionMaterias(matricula=matricula,
#                                                                      agregacion=True,
#                                                                      asignatura=materiaasignada.materia.asignatura,
#                                                                      responsable=persona_responsable,
#                                                                      fecha=datetime.now().date(),
#                                                                      creditos=creditos,
#                                                                      nivelmalla=materiaasignada.materia.nivel.nivelmalla if materiaasignada.materia.nivel.nivelmalla else None,
#                                                                      matriculas=materiaasignada.matriculas)
#                             registro.save()
#                             # if str(matricula.inscripcion_id) not in variable_valor('INSCRIPCIONES_INGLES'):
#                             if matriculas>1 or matricula.inscripcion.persona.tiene_otro_titulo(matricula.inscripcion):
#                                 matricula.nuevo_calculo_matricula_ingles(materiaasignada)
#                             ws.write(row_num, 4, u'%s' % modulomalla, font_style2)
#                             ws.write(row_num, 5, u'%s' % matriculas, font_style2)
#                             ws.write(row_num, 6, u'%s' % materiaasignada, font_style2)
#                             ws.write(row_num, 7, u'%s' % "SI" if matricula.inscripcion.persona.tiene_otro_titulo(matricula.inscripcion) else "NO", font_style2)
#                             row_num += 1
#                             print(u"(%s) -- %s [%s]" % (contador, materiaasignada, materiaasignada.matricula_id))
#                             contador += 1
#                             bandera=1
#                             break
#                         else:
#                             ws.write(row_num, 4, u'YA TIENE MATRICULA EN EL MODULO %s' %materia, font_style2)
#                             ws.write(row_num, 5, u'YA TIENE MATRICULA EN EL MODULO %s' %materia, font_style2)
#                             ws.write(row_num, 6, u'YA TIENE MATRICULA EN EL MODULO %s' %materia, font_style2)
#                             ws.write(row_num, 7, u'YA TIENE MATRICULA EN EL MODULO %s' %materia, font_style2)
#                             row_num += 1
#                             print(u"(%s) -- YA TIENE MATRICULA EN EL MODULO %s" % (contador,materia))
#                             contador += 1
#                             bandera = 1
#             else:
#                 ws.write(row_num, 4, u'NO AY ULTIMO MÓDULO DE INGLÉS PENDIENTE', font_style2)
#                 ws.write(row_num, 5, u'NO AY ULTIMO MÓDULO DE INGLÉS PENDIENTE', font_style2)
#                 ws.write(row_num, 6, u'NO AY ULTIMO MÓDULO DE INGLÉS PENDIENTE', font_style2)
#                 ws.write(row_num, 7, u'NO AY ULTIMO MÓDULO DE INGLÉS PENDIENTE', font_style2)
#                 row_num += 1
#                 print(u"(%s) -- NO AY ULTIMO MÓDULO DE INGLÉS PENDIENTE %s" % (contador,matricula))
#                 contador += 1
#                 bandera = 1
#             # else:
#             #     ws.write(row_num, 4, u'YA TIENE MATRICULA EN AL MENOS UN MODULO %s' %matricula, font_style2)
#             #     ws.write(row_num, 5, u'YA TIENE MATRICULA EN AL MENOS UN MODULO %s' %matricula, font_style2)
#             #     ws.write(row_num, 6, u'YA TIENE MATRICULA EN AL MENOS UN MODULO %s' %matricula, font_style2)
#             #     ws.write(row_num, 7, u'YA TIENE MATRICULA EN AL MENOS UN MODULO %s' %matricula, font_style2)
#             #     row_num += 1
#             #     print(u"(%s) -- YA TIENE MATRICULA EN AL MENOS UN MODULO %s" %(matricula, contador))
#             #     contador += 1
#             #     bandera = 1
#         else:
#             ws.write(row_num, 4, u'NO HAY MODULOS DE INGLES EN LA MALLA', font_style2)
#             ws.write(row_num, 5, u'NO HAY MODULOS DE INGLES EN LA MALLA', font_style2)
#             ws.write(row_num, 6, u'NO HAY MODULOS DE INGLES EN LA MALLA', font_style2)
#             ws.write(row_num, 7, u'NO HAY MODULOS DE INGLES EN LA MALLA', font_style2)
#             row_num += 1
#             print(u"(%s) -- NO HAY MODULOS DE INGLES EN LA MALLA" % contador)
#             contador += 1
#             bandera = 1
#         if bandera == 0:
#             ws.write(row_num, 4, u'NO SE MATRICULA - NO HAY CUPO %s -- %s '%(modulomalla,matricula), font_style2)
#             ws.write(row_num, 5, u'NO SE MATRICULA - NO HAY CUPO %s -- %s '%(modulomalla,matricula), font_style2)
#             ws.write(row_num, 6, u'NO SE MATRICULA - NO HAY CUPO %s -- %s '%(modulomalla,matricula), font_style2)
#             ws.write(row_num, 7, u'NO SE MATRICULA - NO HAY CUPO %s -- %s '%(modulomalla,matricula), font_style2)
#             row_num += 1
#             print(u"NO SE MATRICULA - NO HAY CUPO %s -- %s" % (modulomalla,matricula))
#             contador += 1
#     except Exception as ex:
#         ws.write(row_num, 4, u'SALE ERROR %s' % ( matricula), font_style2)
#         ws.write(row_num, 5, u'SALE ERROR %s' % ( matricula), font_style2)
#         ws.write(row_num, 6, u'SALE ERROR %s' % ( matricula), font_style2)
#         ws.write(row_num, 7, u'SALE ERROR %s' % ( matricula), font_style2)
#         row_num += 1
#         contador += 1
#         print('error: %s %s' % (ex,matricula))
#         pass
# wb.save(filename)
# print("FIN: ", filename)



# miarchivo = openpyxl.load_workbook("insertar_actividades.xlsx")
# lista = miarchivo.get_sheet_by_name('Hoja1')
# totallista = lista.rows
# a=0
# periodo=Periodo.objects.get(id=119)
# for filas in totallista:
#     try:
#         a += 1
#         if a > 1:
#             persona=Persona.objects.filter(id=int(filas[0].value))
#             bitacora = BitacoraActividadDiaria(persona_id=persona,
#                                                fecha=convertir_fecha_hora_invertida(u"%s %s" % (request.POST['fecha'], request.POST['hora'])),
#                                                fechafin=fechahorafin,
#                                                departamento=persona.mi_departamento(),
#                                                descripcion= u'%s' % form.cleaned_data['descripcion'],
#                                                link=form.cleaned_data['link'],
#                                                tiposistema=form.cleaned_data['tiposistema'],
#                                                departamento_requiriente=form.cleaned_data['departamento_requiriente'] )
#     except Exception as ex:
#         print('error: %s' % ex)



def notificar_credenciales_admision(ePeriodo):
    fechaf = (datetime(2023, 10, 25, 0, 0, 0))
    print(f"Inicia proceso de notificación de credenciales para el periodo académico {ePeriodo.__str__()}")
    contador = 0
    eManualUsuario = ManualUsuario.objects.get(id=45)
    miarchivo = openpyxl.load_workbook("cedulas_postulantes_2.xlsx")
    lista = miarchivo.get_sheet_by_name('resultados')
    totallista = lista.rows
    total=16617
    for filas in totallista:
        contador += 1
        if contador>1:
            cedula=str(filas[0].value)
            print(cedula)
            eInscripcion = Inscripcion.objects.filter(Q(persona__cedula=cedula)|Q(persona__pasaporte=cedula),envioemail=False,
                                                        carrera__coordinacion__id=9).first()
            if eInscripcion:
                ePersona = eInscripcion.persona
                lista = []
                if ePersona.email:
                    with transaction.atomic():
                        try:

                            lista.append("mleong2@unemi.edu.ec")
                            lista.append(ePersona.email)
                            send_html_mail("[IMPORTANTE] UNEMI - Credenciales de acceso al Sistema de Gestión Académica",
                                           "emails/email_notificacion_credenciales_admision.html",
                                           {
                                               'sistema': u'Sistema de Gestión Académica',
                                               'fecha': datetime.now().date(),
                                               'hora': datetime.now().time(),
                                               'persona': ePersona,
                                               't': miinstitucion()
                                            },
                                           lista,
                                           [],
                                           [eManualUsuario.archivo],
                                           cuenta=CUENTAS_CORREOS[0][1])
                            print(f"{contador}/{total} -> Correo enviado ({ePersona.email}) +++++ Persona: {ePersona.__str__()} ")
                            eInscripcion.envioemail = True
                            eInscripcion.save()
                            time.sleep(2)
                            # filas[9].value = "CORREO ENVIADO"
                        except Exception as ex:
                            transaction.set_rollback(True)
                            print(f"Ocurrio un error en el envio del correo de la persona: {ePersona.__str__()}")
                else:
                    print(f"{contador}/{total} -> No tiene correo la persona {ePersona.__str__()}")
            else:
                filas[9].value="NO SE ENVIA CORREO"
    print(f"Finalizo proceso de notificación de credenciales para el periodo académico {ePeriodo.__str__()}")
    miarchivo.save("data_cupos_unemi.xlsx")
    print("FIN: ", miarchivo)
#ePeriodo=Periodo.objects.get(id=224)
#notificar_credenciales_admision(ePeriodo)

def notificar_credenciales_admision_2(ePeriodo):
    fechaf = (datetime(2023, 10, 25, 0, 0, 0))
    print(f"Inicia proceso de notificación de credenciales para el periodo académico {ePeriodo.__str__()}")
    contador = 0
    eManualUsuario = ManualUsuario.objects.get(id=45)
    matriculas = Matricula.objects.filter(status=True, nivel__id__in=[1516,1517], nivel__periodo__id=224, inscripcion__envioemail=False, termino=True, fechatermino__isnull=True, cuposenescyt=True)
    total=matriculas.count()
    for matri in matriculas:
        contador += 1
        if contador>1:
            eInscripcion=matri.inscripcion
            persona= matri.inscripcion.persona
            lista = []
            if persona.email:
                with transaction.atomic():
                    try:
                        lista.append(persona.email)
                        send_html_mail("[IMPORTANTE] UNEMI - Credenciales de acceso al Sistema de Gestión Académica",
                                       "emails/email_notificacion_credenciales_admision.html",
                                       {
                                           'sistema': u'Sistema de Gestión Académica',
                                           'fecha': datetime.now().date(),
                                           'hora': datetime.now().time(),
                                           'persona': persona,
                                           't': miinstitucion()
                                       },
                                       lista,
                                       [],
                                       [eManualUsuario.archivo],
                                       cuenta=CUENTAS_CORREOS[0][1])
                        print(
                            f"{contador}/{total} -> Correo enviado ({persona.email}) +++++ Persona: {persona.__str__()} ")
                        eInscripcion.envioemail = True
                        eInscripcion.save()
                        time.sleep(2)


                            # filas[9].value = "CORREO ENVIADO"
                    except Exception as ex:
                        transaction.set_rollback(True)
                        print(f"Ocurrio un error en el envio del correo de la persona: {persona.__str__()}")

            else:
                print(f"{contador}/{total} -> No tiene correo la persona {persona.__str__()}")

    print(f"Finalizo proceso de notificación de credenciales para el periodo académico {ePeriodo.__str__()}")

# ePeriodo=Periodo.objects.get(id=224)
# notificar_credenciales_admision_2(ePeriodo)


# for materiaasignada in MateriaAsignada.objects.filter(status=True, matricula=matricula, retiramateria=False):
#     materiaasignada.importa_nota = True
#     materiaasignada.cerrado = True
#     materiaasignada.fechacierre = datetime(2022, 10, 6, 0, 0, 0).date()
#     materiaasignada.save()
#     d = locals()
#     exec(materiaasignada.materia.modeloevaluativo.logicamodelo, globals(), d)
#     d['calculo_modelo_evaluativo'](materiaasignada)
#     materiaasignada.cierre_materia_asignada()
#     print(u"CIERRA -- %s" % (materiaasignada))

# fechaf = (datetime(2022, 11, 18, 0, 0, 0))
# matriculas=Matricula.objects.filter(status=True,nivel__periodo_id=202, fecha_creacion__gte=fechaf)
# print(matriculas.count())
# contador=0
# for matricula in matriculas:
#     oPersona=matricula.inscripcion.persona
#     if not oPersona.es_administrativo() and not oPersona.es_profesor():
#         if not oPersona.necesita_cambiar_clave():
#             oUser = oPersona.usuario
#             oUser.is_active = True
#             oUser.set_password(oPersona.identificacion())
#             oUser.save()
#             oPersona.cambiar_clave()
#             print(matricula)

# def delete_materiaasignada():
#     with transaction.atomic():
#         try:
#             alumnos = MateriaAsignada.objects.filter(status=True,
#                                                      matricula__inscripcion__persona__cedula__in=['0917989873',
#                                                                                                   '1004144653',
#                                                                                                   '1206119818',
#                                                                                                   '2450173956',
#                                                                                                   '0750478661',
#                                                                                                   '0926739210',
#                                                                                                   '0927339168',
#                                                                                                   '1206323576',
#                                                                                                   '1206715771',
#                                                                                                   '0919014886',
#                                                                                                   '0705030724',
#                                                                                                   '0602685968',
#                                                                                                   '0925273344',
#                                                                                                   '1208663870',
#                                                                                                   '0950223503',
#                                                                                                   '1718781295',
#                                                                                                   '2200077390',
#                                                                                                   '2150080832',
#                                                                                                   '0931133979',
#                                                                                                   '0801247081',
#                                                                                                   '0941046369',
#                                                                                                   '0917511990'
#                                                                                                   ],
#                                                      materia_id=59155)
#             print(len(alumnos))
#             for alumno in alumnos:
#                 print(alumno.matricula.inscripcion.persona, alumno.materia)
#                 alumno.delete()
#
#         except Exception as e:
#             transaction.set_rollback(True)
#             print('error: %s' % e)


# delete_materiaasignada()
# eVariable = VariablesGlobales.objects.get(pk=105)
# eVariable.valor="False"
# eVariable.save()


def reporte_clicks():
    response = HttpResponse(content_type="application/ms-excel")
    response['Content-Disposition'] = 'attachment; filename=reporteclics.xls'
    style0 = easyxf('font: name Times New Roman, color-index blue, bold off', num_format_str='#,##0.00')
    style_nb = easyxf('font: name Times New Roman, color-index blue, bold on', num_format_str='#,##0.00')
    style_sb = easyxf('font: name Times New Roman, color-index blue, bold on')
    title = easyxf('font: name Times New Roman, color-index blue, bold on , height 350; alignment: horiz centre')
    style1 = easyxf(num_format_str='D-MMM-YY')
    font_style = XFStyle()
    font_style.font.bold = True
    font_style2 = XFStyle()
    font_style2.font.bold = False
    wb = xlwt.Workbook()
    ws = wb.add_sheet('Sheetname')
    estilo = xlwt.easyxf('font: height 350, name Arial, colour_index black, bold on, italic on; align: wrap on, vert centre, horiz center;')
    ws.write_merge(0, 0, 0, 9, 'UNIVERSIDAD ESTATAL DE MILAGRO', estilo)
    output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
    nombre = "Clics.xls"
    filename = os.path.join(output_folder, nombre)
    columns = [(u"url", 6000),
               (u"modulo", 6000),
               (u"opcion", 6000),
               (u"cedula", 6000),
               (u"sexo", 6000),
               (u"fecha nacimiento", 6000),
               (u"facultad", 6000),
               (u"carrera", 6000),
               (u"fecha creacion", 6000)
               ]
    row_num = 3
    for col_num in range(len(columns)):
        ws.write(row_num, col_num, columns[col_num][0], font_style)
        ws.col(col_num).width = columns[col_num][1]
    row_num = 4
    conexion = connections['uxplora']
    cnunemi = conexion.cursor()
    sql1 = f" select module,ACTION,userid,inscripcion,url,fecha_creacion from logstore_logstoresystem limit 5  "
    cnunemi.execute(sql1)
    registros = cnunemi.fetchall()

    for gdata in registros:
        modulo=Modulo.objects.get(id=gdata[0])
        opcion = InventarioOpcionSistema.objects.get(id=gdata[1])
        persona=Persona.objects.get(usuario_id=gdata[2])
        inscripcion=Inscripcion.objects.get(id=gdata[3])
        ws.write(row_num, 0, u'%s' % gdata.url, font_style2)
        ws.write(row_num, 1, u'%s' % modulo if modulo else '', font_style2)
        ws.write(row_num, 2, u'%s' % opcion if opcion else '', font_style2)
        ws.write(row_num, 3, u'%s' % persona.identificacion() if persona else '', font_style2)
        ws.write(row_num, 4, u'%s' % persona.sexo_id if persona else '', font_style2)
        ws.write(row_num, 5, u'%s' % persona.nacimiento if persona else '', font_style2)
        ws.write(row_num, 6, u'%s' % inscripcion.carrera.coordinacion_carrera().nombre if inscripcion else '', font_style2)
        ws.write(row_num, 7, u'%s' % inscripcion.carrera if inscripcion else '', font_style2)
        ws.write(row_num, 8, u'%s' % gdata[4].strftime('%Y-%m-%d %H:%M:%S'), font_style2)
        print(u"%s" % gdata)
        row_num += 1
    wb.save(filename)
    print("FIN: ", filename)




# materia = Materia.objects.get(pk=68220, status=True)
# estudiantes = materia.asignados_a_esta_materia_moodle().filter(retiramateria=False)
# primerestudiante = estudiantes.filter(matricula__bloqueomatricula=False).first()
# print(primerestudiante)
# bandera = True
# modelo_mood = ''
# modelo_sga = ''
# for notassga in primerestudiante.evaluacion_generica():
#     modelo_sga += "{}, ".format(notassga.detallemodeloevaluativo.nombre)
# print(modelo_mood)

# for notasmooc in materia.notas_de_moodle(primerestudiante.matricula.inscripcion.persona):
#     print(notasmooc[1].upper().strip())
#     bandera = primerestudiante.evaluacion_generica().filter(detallemodeloevaluativo__nombre=notasmooc[1].upper().strip()).exists()
#     if not bandera:
#         for notasmoocstr in materia.notas_de_moodle(primerestudiante.matricula.inscripcion.persona):
#             modelo_mood += "{}, ".format(notasmoocstr[1])
#         for notassga in primerestudiante.evaluacion_generica():
#             modelo_sga += "{}, ".format(notassga.detallemodeloevaluativo.nombre)
#         print(modelo_mood)
#         print(modelo_sga)


def reporte_examen_pregrado_en_linea(ePeriodo,fecha):
    try:
        libre_origen = '/reporte_examen_sede_moodle_pregrado_jussy_v2'+ datetime.now().strftime('%Y%m%d_%H%M%S') +'.xls'
        fuentecabecera = easyxf('font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
        fuentenormal = easyxf('font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
        # output_folder = MEDIA_ROOT
        output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
        # output_folder = os.path.join(os.path.join(BASE_DIR))
        # liborigen = xlrd.open_workbook(output_folder + libre_origen)
        libdestino = xlwt.Workbook()
        hojadestino = libdestino.add_sheet('HOJA1')
        fil = 0
        columnas = [
            (u"facultad", 7000, 1),
            (u"carrera", 7000, 1),
            (u"nivel", 7000, 0),
            (u"paralelo", 7000, 0),
            (u"asignatura", 7000, 0),
            (u"docente", 7000, 0),
            (u"idmateria", 7000, 0),
            (u"idcursomoodle", 7000, 0),
            (u"examen_creado_sga", 7000, 0),
            (u"modelo_evaluativo_sga", 7000, 0),
            (u"migrado_moodle", 7000, 0),
            (u"nombre_examen_moodle", 7000, 0),
            (u"desde_examen_moodle", 7000, 0),
            (u"hasta_examen_moodle", 7000, 0),
            (u"tiempo_examen_moodle", 7000, 0),
            (u"metodo_navegacion_examen_moodle", 7000, 0),
            (u"total_examen_moodle", 7000, 0),
            (u"valor_examen_moodle", 7000, 0),
            (u"tiene_clave_examen_moodle", 7000, 0),
            (u"categoria_calificacion_examen_moodle", 7000, 0),
            (u"item_calificacion_examen_moodle", 7000, 0),
            (u"seccion_examen_moodle", 7000, 0),
            (u"tiene_preguntas", 7000, 0),
            (u"nombre examen general", 7000, 0),
            (u"id examen general", 7000, 0),
            (u"id examen silabo", 7000, 0),
        ]
        for col_num in range(len(columnas)):
            hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
            hojadestino.col(col_num).width = columnas[col_num][1]
        id_materias = MateriaAsignadaPlanificacionSedeVirtualExamen.objects.values_list('materiaasignada__materia_id').filter(status=True,
                                                  materiaasignada__materia__nivel__periodo=ePeriodo,
                                                  aulaplanificacion__turnoplanificacion__fechaplanificacion__fecha__range=(fecha),
                                                  ).distinct()
        # eMaterias = Materia.objects.filter(Q(asignaturamalla__malla__carrera__modalidad=3) | Q(asignaturamalla__asignatura__id__in=asignaturas_id), status=True, nivel__periodo=ePeriodo).exclude(nivel__id__in=[1481, 1482, 1501, 1508])
        eMaterias = Materia.objects.filter(status=True,nivel__periodo=ePeriodo,id__in=id_materias ).exclude(nivel__id__in=[1481, 1482, 1501, 1508,1542])
        totalmaterias = eMaterias.count()
        print(totalmaterias)
        cont = 1
        fila = 1
        for eMateria in eMaterias:
            facultad = eMateria.coordinacion()
            carrera = eMateria.asignaturamalla.malla.carrera
            nivel = eMateria.asignaturamalla.nivelmalla
            paralelo = eMateria.paralelo
            asignatura = eMateria.asignaturamalla.asignatura.nombre
            profesor = 'sin profesor'
            pm = eMateria.profesormateria_set.filter(status=True, tipoprofesor_id=14).last()
            if pm:
                profesor = pm.profesor.persona.nombre_completo_inverso()
            idmateria = eMateria.id
            idcurso = eMateria.idcursomoodle
            tiene_examen_sga= 'NO'
            migrado_moodle= 'NO'
            num_preguntas = 0
            modeloevaluativo = 'SIN MODELO'
            if eMateria.modeloevaluativo:
                modeloevaluativo = eMateria.modeloevaluativo.nombre
                detallemodelo_evaluativo_examen=DetalleModeloEvaluativo.objects.filter(status=True,migrarmoodle=True,alternativa_id=20,modelo=eMateria.modeloevaluativo).order_by('orden').last()
            eTestSilaboSemanal = TestSilaboSemanal.objects.filter(status=True, silabosemanal__silabo__materia_id=idmateria, silabosemanal__examen=True, detallemodelo__alternativa_id=20).last()
            if facultad.id == 9:
                cursor_verbose = 'db_moodle_virtual'
            else:
                cursor_verbose = 'moodle_db'
            conexion = connections[cursor_verbose]
            cursor = conexion.cursor()
            name = ''
            timeopen = ''
            timeclose = ''
            timelimit = ''
            navmethod = ''
            sumgrades = ''
            grade = ''
            password = None
            categoria = eTestSilaboSemanal.detallemodelo.nombre if eTestSilaboSemanal else ''
            categoryid = 0
            categoria_moodle = ''
            itemid = 0
            item_moodle = ''
            seccion = ''
            instance = None
            if eTestSilaboSemanal:
                tiene_examen_sga = 'SI'
                if eTestSilaboSemanal.estado_id == 4:
                    migrado_moodle = 'SI'
                    instance = eTestSilaboSemanal.idtestmoodle
                if instance:
                    sql = f"""SELECT name, timeopen, timeclose, timelimit, navmethod, sumgrades, grade, password  FROM mooc_quiz WHERE id={instance} AND course={eMateria.idcursomoodle}"""
                    cursor.execute(sql)
                    quiz = cursor.fetchone()
                    if quiz:
                        name = quiz[0]
                        timeopen = quiz[1]
                        timeopen = str(datetime.fromtimestamp(timeopen))
                        timeclose = quiz[2]
                        timeclose = str(datetime.fromtimestamp(timeclose))
                        timelimit = quiz[3]
                        timelimit = str((timelimit / 60) if timelimit else 0)
                        navmethod = quiz[4]
                        sumgrades = str(quiz[5])
                        grade = str(quiz[6])
                        password = str(quiz[7])
                        sql = """SELECT id, fullname FROM mooc_grade_categories WHERE courseid=%s AND fullname='%s' and depth='2' """ % (eMateria.idcursomoodle, categoria)
                        cursor.execute(sql)
                        category = cursor.fetchone()
                        if category:
                            categoryid = category[0]
                            categoria_moodle = category[1]
                            sql = """select id, itemname from mooc_grade_items WHERE courseid=%s AND categoryid=%s and itemname='%s' and iteminstance=%s """ % (eMateria.idcursomoodle, categoryid, name, instance)
                            cursor.execute(sql)
                            item = cursor.fetchone()
                            if item:
                                itemid = item[0]
                                item_moodle = item[1]
                        sql = """SELECT section FROM mooc_course_modules WHERE course=%s AND instance=%s """ % (eMateria.idcursomoodle, instance)
                        cursor.execute(sql)
                        course_module = cursor.fetchone()
                        if course_module:
                            sectionid = course_module[0]
                            sql = """SELECT name FROM mooc_course_sections WHERE course=%s AND id=%s """ % (eMateria.idcursomoodle, sectionid)
                            cursor.execute(sql)
                            section = cursor.fetchone()
                            if section:
                                seccion = section[0]
                        sql = """SELECT DISTINCT  qet.name, qet.questiontext, re.answer, re.answerformat FROM 
                        mooc_quiz q INNER JOIN mooc_quiz_slots qe ON q.id=qe.quizid INNER JOIN mooc_question qet ON 
                        qet.category=qe.questioncategoryid INNER JOIN mooc_question_answers re ON re.question=qet.id 
                        WHERE re.fraction>0 AND q.id=%s  """ % (instance)
                        cursor.execute(sql)
                        preguntas = cursor.fetchall()
                        num_preguntas = len(preguntas)
            name_1=''
            quizid_1=''
            if detallemodelo_evaluativo_examen:
                sql_quiz = u"""
                        select quiz.name, quiz.timeopen, quiz.timeclose, quiz.timelimit, quiz.navmethod, quiz.sumgrades, 
                        quiz.grade, quiz.password, quiz.id
                        from mooc_grade_items gi 
                        inner join mooc_grade_categories gc on gi.courseid=gc.courseid and gc.id=gi.CATEGORYID
                        inner join mooc_quiz quiz on quiz.id=gi.ITEMINSTANCE
                        where gc.courseid=%s and gc.FULLNAME='%s'"""%(eMateria.idcursomoodle,detallemodelo_evaluativo_examen.nombre)
                cursor.execute(sql_quiz)
                quiz_moodle = cursor.fetchone()
                if quiz_moodle:
                    name_1 = quiz_moodle[0]
                    quizid_1 = quiz_moodle[8]
            print('curso actualizado', cont, 'de', totalmaterias)
            cont += 1

            hojadestino.write(fila, 0, "%s" % facultad, fuentenormal)
            hojadestino.write(fila, 1, "%s" % carrera, fuentenormal)
            hojadestino.write(fila, 2, "%s" % nivel, fuentenormal)
            hojadestino.write(fila, 3, "%s" % paralelo, fuentenormal)
            hojadestino.write(fila, 4, "%s" % asignatura, fuentenormal)
            hojadestino.write(fila, 5, "%s" % profesor, fuentenormal)
            hojadestino.write(fila, 6, idmateria, fuentenormal)
            hojadestino.write(fila, 7, idcurso, fuentenormal)
            hojadestino.write(fila, 8, tiene_examen_sga, fuentenormal)
            hojadestino.write(fila, 9, modeloevaluativo, fuentenormal)
            hojadestino.write(fila, 10, migrado_moodle, fuentenormal)
            hojadestino.write(fila, 11, name, fuentenormal)
            hojadestino.write(fila, 12, timeopen, fuentenormal)
            hojadestino.write(fila, 13, timeclose, fuentenormal)
            hojadestino.write(fila, 14, timelimit, fuentenormal)
            hojadestino.write(fila, 15, navmethod, fuentenormal)
            hojadestino.write(fila, 16, sumgrades, fuentenormal)
            hojadestino.write(fila, 17, grade, fuentenormal)
            hojadestino.write(fila, 18, 'SI' if password else 'NO', fuentenormal)
            hojadestino.write(fila, 19, categoria_moodle, fuentenormal)
            hojadestino.write(fila, 20, item_moodle, fuentenormal)
            hojadestino.write(fila, 21, seccion, fuentenormal)
            hojadestino.write(fila, 22, 'SI' if num_preguntas > 0 else 'NO', fuentenormal)
            hojadestino.write(fila, 23, '%s'%name_1, fuentenormal)
            hojadestino.write(fila, 24, '%s'%quizid_1, fuentenormal)
            hojadestino.write(fila, 25, '%s'%instance, fuentenormal)

            fila += 1

        libdestino.save(output_folder + libre_origen)
        print(output_folder + libre_origen)
        print("Proceso finalizado. . .")

    except Exception as ex:
        msg = ex.__str__()
        textoerror = '{} Linea:{}'.format(str(ex), sys.exc_info()[-1].tb_lineno)
        print(textoerror)
        print(msg)


def reporte_examen_pregrado_en_linea_detalle(ePeriodo,fecha):
    try:
        libre_origen = '/reporte_examen_sede_moodle_pregrado_jussy_detalle'+ datetime.now().strftime('%Y%m%d_%H%M%S') +'.xls'
        fuentecabecera = easyxf('font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
        fuentenormal = easyxf('font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
        # output_folder = MEDIA_ROOT
        output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media'))
        # output_folder = os.path.join(os.path.join(BASE_DIR))
        # liborigen = xlrd.open_workbook(output_folder + libre_origen)
        libdestino = xlwt.Workbook()
        hojadestino = libdestino.add_sheet('HOJA1')
        fil = 0
        columnas = [
            (u"facultad", 7000, 1),
            (u"carrera", 7000, 1),
            (u"nivel", 7000, 0),
            (u"paralelo", 7000, 0),
            (u"asignatura", 7000, 0),
            (u"docente", 7000, 0),
            (u"idmateria", 7000, 0),
            (u"idcursomoodle", 7000, 0),
            (u"examen_creado_sga", 7000, 0),
            (u"modelo_evaluativo_sga", 7000, 0),
            (u"migrado_moodle", 7000, 0),
            (u"nombre_examen_moodle", 7000, 0),
            (u"desde_examen_moodle", 7000, 0),
            (u"hasta_examen_moodle", 7000, 0),
            (u"tiempo_examen_moodle", 7000, 0),
            (u"metodo_navegacion_examen_moodle", 7000, 0),
            (u"total_examen_moodle", 7000, 0),
            (u"valor_examen_moodle", 7000, 0),
            (u"tiene_clave_examen_moodle", 7000, 0),
            (u"categoria_calificacion_examen_moodle", 7000, 0),
            (u"item_calificacion_examen_moodle", 7000, 0),
            (u"seccion_examen_moodle", 7000, 0),
            (u"tiene_preguntas", 7000, 0),
            (u"examen_solo", 7000, 0),
            (u"id examen solo", 7000, 0),
            (u"id examen planificacion", 7000, 0),
            (u"id planificacion", 7000, 0),
            (u"id SILABO SEMANAL", 7000, 0),
            (u"ALUMNO", 7000, 0)
        ]
        for col_num in range(len(columnas)):
            hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
            hojadestino.col(col_num).width = columnas[col_num][1]
        planificacion = MateriaAsignadaPlanificacionSedeVirtualExamen.objects.filter(status=True,
                                                  materiaasignada__materia__nivel__periodo=ePeriodo,
                                                  materiaasignada____asignaturamalla__malla__carrera__coordinacion__in=[1,2,3,4,5],
                                                  aulaplanificacion__turnoplanificacion__fechaplanificacion__fecha__range=(fecha),
                                                  ).distinct()
        cont = 1
        fila = 1
        totalmaterias=planificacion.count()
        for eplanificacion in planificacion:
            facultad = eplanificacion.materiaasignada.materia.coordinacion()
            carrera = eplanificacion.materiaasignada.materia.asignaturamalla.malla.carrera
            nivel = eplanificacion.materiaasignada.materia.asignaturamalla.nivelmalla
            paralelo = eplanificacion.materiaasignada.materia.paralelo
            asignatura = eplanificacion.materiaasignada.materia.asignaturamalla.asignatura.nombre
            profesor = 'sin profesor'
            pm = eplanificacion.materiaasignada.materia.profesormateria_set.filter(status=True, tipoprofesor_id=14).last()
            if pm:
                profesor = pm.profesor.persona.nombre_completo_inverso()
            idmateria = eplanificacion.materiaasignada.materia.id
            idcurso = eplanificacion.materiaasignada.materia.idcursomoodle
            tiene_examen_sga= 'NO'
            migrado_moodle= 'NO'
            num_preguntas = 0
            modeloevaluativo = 'SIN MODELO'
            if eplanificacion.materiaasignada.materia.modeloevaluativo:
                modeloevaluativo = eplanificacion.materiaasignada.materia.modeloevaluativo.nombre
                detallemodelo_evaluativo_examen=DetalleModeloEvaluativo.objects.filter(status=True,migrarmoodle=True,alternativa_id=20,modelo=eplanificacion.materiaasignada.materia.modeloevaluativo).order_by('orden').last()
            eTestSilaboSemanal = TestSilaboSemanal.objects.filter(status=True, silabosemanal__silabo__materia_id=idmateria, silabosemanal__examen=True, detallemodelo__alternativa_id=20).last()
            if facultad.id == 9:
                cursor_verbose = 'db_moodle_virtual'
            else:
                cursor_verbose = 'moodle_db'
            conexion = connections[cursor_verbose]
            cursor = conexion.cursor()
            name = ''
            timeopen = ''
            timeclose = ''
            timelimit = ''
            navmethod = ''
            sumgrades = ''
            grade = ''
            password = None
            categoria = eTestSilaboSemanal.detallemodelo.nombre if eTestSilaboSemanal else ''
            categoryid = 0
            categoria_moodle = ''
            itemid = 0
            item_moodle = ''
            seccion = ''
            instance = None
            if eTestSilaboSemanal:
                tiene_examen_sga = 'SI'
                if eTestSilaboSemanal.estado_id == 4:
                    migrado_moodle = 'SI'
                    instance = eTestSilaboSemanal.idtestmoodle
                if instance:
                    sql = f"""SELECT name, timeopen, timeclose, timelimit, navmethod, sumgrades, grade, password  FROM mooc_quiz WHERE id={instance} AND course={eplanificacion.materiaasignada.materia.idcursomoodle}"""
                    cursor.execute(sql)
                    quiz = cursor.fetchone()
                    if quiz:
                        name = quiz[0]
                        timeopen = quiz[1]
                        timeopen = str(datetime.fromtimestamp(timeopen))
                        timeclose = quiz[2]
                        timeclose = str(datetime.fromtimestamp(timeclose))
                        timelimit = quiz[3]
                        timelimit = str((timelimit / 60) if timelimit else 0)
                        navmethod = quiz[4]
                        sumgrades = str(quiz[5])
                        grade = str(quiz[6])
                        password = str(quiz[7])
                        sql = """SELECT id, fullname FROM mooc_grade_categories WHERE courseid=%s AND fullname='%s' and depth='2' """ % (eplanificacion.materiaasignada.materia.idcursomoodle, categoria)
                        cursor.execute(sql)
                        category = cursor.fetchone()
                        if category:
                            categoryid = category[0]
                            categoria_moodle = category[1]
                            sql = """select id, itemname from mooc_grade_items WHERE courseid=%s AND categoryid=%s and itemname='%s' and iteminstance=%s """ % (eplanificacion.materiaasignada.materia.idcursomoodle, categoryid, name, instance)
                            cursor.execute(sql)
                            item = cursor.fetchone()
                            if item:
                                itemid = item[0]
                                item_moodle = item[1]
                        sql = """SELECT section FROM mooc_course_modules WHERE course=%s AND instance=%s """ % (eplanificacion.materiaasignada.materia.idcursomoodle, instance)
                        cursor.execute(sql)
                        course_module = cursor.fetchone()
                        if course_module:
                            sectionid = course_module[0]
                            sql = """SELECT name FROM mooc_course_sections WHERE course=%s AND id=%s """ % (eplanificacion.materiaasignada.materia.idcursomoodle, sectionid)
                            cursor.execute(sql)
                            section = cursor.fetchone()
                            if section:
                                seccion = section[0]
                        sql = """SELECT DISTINCT  qet.name, qet.questiontext, re.answer, re.answerformat FROM 
                        mooc_quiz q INNER JOIN mooc_quiz_slots qe ON q.id=qe.quizid INNER JOIN mooc_question qet ON 
                        qet.category=qe.questioncategoryid INNER JOIN mooc_question_answers re ON re.question=qet.id 
                        WHERE re.fraction>0 AND q.id=%s  """ % (instance)
                        cursor.execute(sql)
                        preguntas = cursor.fetchall()
                        num_preguntas = len(preguntas)
            name_1=''
            quizid_1=''
            if detallemodelo_evaluativo_examen:
                sql_quiz = u"""
                        select quiz.name, quiz.timeopen, quiz.timeclose, quiz.timelimit, quiz.navmethod, quiz.sumgrades, 
                        quiz.grade, quiz.password, quiz.id
                        from mooc_grade_items gi 
                        inner join mooc_grade_categories gc on gi.courseid=gc.courseid and gc.id=gi.CATEGORYID
                        inner join mooc_quiz quiz on quiz.id=gi.ITEMINSTANCE
                        where gc.courseid=%s and gc.FULLNAME='%s'"""%(eplanificacion.materiaasignada.materia.idcursomoodle,detallemodelo_evaluativo_examen.nombre)
                cursor.execute(sql_quiz)
                quiz_moodle = cursor.fetchone()
                if quiz_moodle:
                    name_1 = quiz_moodle[0]
                    quizid_1 = quiz_moodle[8]
            print('EVALUADO', cont, 'de', totalmaterias)
            cont += 1

            hojadestino.write(fila, 0, "%s" % facultad, fuentenormal)
            hojadestino.write(fila, 1, "%s" % carrera, fuentenormal)
            hojadestino.write(fila, 2, "%s" % nivel, fuentenormal)
            hojadestino.write(fila, 3, "%s" % paralelo, fuentenormal)
            hojadestino.write(fila, 4, "%s" % asignatura, fuentenormal)
            hojadestino.write(fila, 5, "%s" % profesor, fuentenormal)
            hojadestino.write(fila, 6, idmateria, fuentenormal)
            hojadestino.write(fila, 7, idcurso, fuentenormal)
            hojadestino.write(fila, 8, tiene_examen_sga, fuentenormal)
            hojadestino.write(fila, 9, modeloevaluativo, fuentenormal)
            hojadestino.write(fila, 10, migrado_moodle, fuentenormal)
            hojadestino.write(fila, 11, name, fuentenormal)
            hojadestino.write(fila, 12, timeopen, fuentenormal)
            hojadestino.write(fila, 13, timeclose, fuentenormal)
            hojadestino.write(fila, 14, timelimit, fuentenormal)
            hojadestino.write(fila, 15, navmethod, fuentenormal)
            hojadestino.write(fila, 16, sumgrades, fuentenormal)
            hojadestino.write(fila, 17, grade, fuentenormal)
            hojadestino.write(fila, 18, 'SI' if password else 'NO', fuentenormal)
            hojadestino.write(fila, 19, categoria_moodle, fuentenormal)
            hojadestino.write(fila, 20, item_moodle, fuentenormal)
            hojadestino.write(fila, 21, seccion, fuentenormal)
            hojadestino.write(fila, 22, 'SI' if num_preguntas > 0 else 'NO', fuentenormal)
            hojadestino.write(fila, 23, '%s'%name_1, fuentenormal)
            hojadestino.write(fila, 24, '%s'%quizid_1, fuentenormal)
            hojadestino.write(fila, 25, '%s'%eplanificacion.idtestmoodle, fuentenormal)
            hojadestino.write(fila, 26, '%s'%eplanificacion.id, fuentenormal)
            hojadestino.write(fila, 27, '%s'%instance, fuentenormal)
            hojadestino.write(fila, 28, '%s'%eplanificacion, fuentenormal)
            fila += 1
        libdestino.save(output_folder + libre_origen)
        print(output_folder + libre_origen)
        print("Proceso finalizado. . .")

    except Exception as ex:
        msg = ex.__str__()
        textoerror = '{} Linea:{}'.format(str(ex), sys.exc_info()[-1].tb_lineno)
        print(textoerror)
        print(msg)



# reporte_examen_pregrado_en_linea(Periodo.objects.get(id=224),[convertir_fecha('06-01-2024'), convertir_fecha('06-01-2024')])


def agregarinsignia():
    ruta = os.path.join(SITE_ROOT, 'runback', 'arreglos', 'media')
    excel = openpyxl.load_workbook("{}{}maestrias2023.xlsx".format(ruta, os.sep))
    insignia = Insignia.objects.get(id=5, status=True)
    listaNoexiste = []
    listaInsigniaExiste = []
    cont = 0
    cont2 = 0
    cont3 = 0
    for i in excel.sheetnames[0:]:
        lista = excel[i]
        numlist = lista.rows
        cab = 0
        for personaIn in numlist:
            cab += 1
            personaI = None
            if cab > 1:
                if Persona.objects.filter(cedula=personaIn[0].value).order_by('-id').exists():
                    personaInsign = Persona.objects.filter(cedula=personaIn[0].value).order_by('-id').first()
                    if not InsigniaPersona.objects.filter(persona=personaInsign, insignia=insignia,
                                                          status=True).order_by('-id').exists():
                        insiniapersonanueva = InsigniaPersona(
                            persona=personaInsign,
                            insignia=insignia,
                            fechaobtencion=personaIn[3].value
                        )
                        insiniapersonanueva.save()
                        cont += 1
                        print(f'No. {cont} {insiniapersonanueva} - {personaInsign} - {insignia}')
                    else:
                        cont2 += 1
                        listaInsigniaExiste.append({'cedula': personaIn[0].value, 'nombre': personaIn[1].value})
                else:
                    cont3 += 1
                    listaNoexiste.append({'cedula': personaIn[0].value, 'nombre': personaIn[1].value})
    print(f'Total de registro ingresados: {cont}')
    print(f'Total de personas que no se encontraron ({cont3}) {listaNoexiste}')
    print(f'Total de personas que ya tienen la insignia({cont2}) {listaInsigniaExiste}')


# agregarinsignia()


def agregarinsigniaxnombres():
    ruta = os.path.join(SITE_ROOT,  'media')
    excel = openpyxl.load_workbook("{}{}MEJORES EVALUADOS-11-12-23.xlsx".format(ruta, os.sep))
    insignia = Insignia.objects.get(id=16, status=True)
    listaNoexiste = []
    listaInsigniaExiste = []
    cont = 0
    cont2 = 0
    cont3 = 0
    for i in excel.sheetnames[0:]:
        lista = excel[i]
        numlist = lista.rows
        cab = 0
        for personaIn in numlist:
            cab += 1
            personaI = None
            if cab > 1:
                cadena=personaIn[0].value
                dividido=cadena.split()
                apellido1=dividido[0]
                apellido2=dividido[1]
                nombre1=dividido[2]
                if Persona.objects.filter(apellido1=apellido1,apellido2=apellido2).order_by('-id').exists():
                    personaInsign = Persona.objects.filter(apellido1=apellido1,apellido2=apellido2, nombres__icontains=nombre1).order_by('-id').first()
                    if not InsigniaPersona.objects.filter(persona=personaInsign, insignia=insignia,
                                                          status=True).order_by('-id').exists():
                        insiniapersonanueva = InsigniaPersona(
                            persona=personaInsign,
                            insignia=insignia,
                            fechaobtencion=personaIn[1].value
                        )
                        insiniapersonanueva.save()
                        cont += 1
                        print(f'No. {cont} {insiniapersonanueva} - {personaInsign} - {insignia}')
                    else:
                        cont2 += 1
                        listaInsigniaExiste.append({'cedula': personaIn[0].value, 'nombre': personaIn[1].value})
                else:
                    cont3 += 1
                    listaNoexiste.append({'cedula': personaIn[0].value, 'nombre': personaIn[1].value})
    print(f'Total de registro ingresados: {cont}')
    print(f'Total de personas que no se encontraron ({cont3}) {listaNoexiste}')
    print(f'Total de personas que ya tienen la insignia({cont2}) {listaInsigniaExiste}')


# agregarinsigniaxnombres()


def inscribir_libre_oma(cedula,curso):
    if Persona.objects.filter(Q(cedula=cedula) or Q(pasaporte=cedula)).exists():
        personainsc = Persona.objects.filter(Q(cedula=cedula) or Q(pasaporte=cedula)).first()
        correo = personainsc.email
        insccurso = None
        try:
            insccurso = InscripcionCurso.objects.get(persona=personainsc, curso=curso,
                                                     status=True)
        except ObjectDoesNotExist:
            insccurso = InscripcionCurso(persona=personainsc, curso=curso, correo=correo)
            insccurso.save()
    for y in DetalleModeloEvaluativo.objects.filter(modelo=curso.modeloevaluativo, status=True):
        try:
            evaluagenerica = EvaluacionGenerica.objects.get(detallemodeloevaluativo=y,
                                                            inscripcioncurso=insccurso,
                                                            status=True)
        except ObjectDoesNotExist:
            evaluagenerica = EvaluacionGenerica(detallemodeloevaluativo=y,
                                                inscripcioncurso=insccurso)
            evaluagenerica.save()

inscribir_libre_oma('0941689705',Curso.objects.get(id=12))
# inscribir_libre_oma('0941689705',Curso.objects.get(id=12))

def jefrersion():
    from sga.funciones import notificacion2
    persona = Persona.objects.get(id=37121)
    try:
        periodo = Periodo.objects.get(id=317)
        response_str = 'Listado de docentes afectados: '
        for i, detalle in enumerate(DetalleDistributivo.objects.filter(distributivo__periodo=periodo,
                                                                       criterioinvestigacionperiodo__criterio__id=68,
                                                                       status=True)):
            if actividad := detalle.actividaddetalledistributivo_set.filter(status=True, vigente=True).first():
                for subactividad in SubactividadDocentePeriodo.objects.filter(
                        actividad__criterioinvestigacionperiodo=detalle.criterioinvestigacionperiodo,
                        actividad__status=True, status=True):
                    if not SubactividadDetalleDistributivo.objects.values('id').filter(
                            actividaddetalledistributivo=actividad, subactividaddocenteperiodo=subactividad,
                            status=True).exists():
                        s = SubactividadDetalleDistributivo(actividaddetalledistributivo=actividad,
                                                            subactividaddocenteperiodo=subactividad,
                                                            fechainicio=periodo.inicio, fechafin=periodo.fin)
                        s.save()
            response_str += f"""
                <br>{i}, {detalle.pk}, {detalle.distributivo.profesor.persona.__str__()}
            """
        notificacion2("Resultados arreglo jefferson", response_str, persona, None, '/notificacion', persona.pk, 1,
                      'sga', Persona)
        DEBUG and print(response_str)
    except Exception as ex:
        linea_error = 'Error on line {}'.format(sys.exc_info()[-1].tb_lineno)
        notificacion2("Resultados arreglo jefferson", f"{linea_error}. {ex.__str__()}", persona, None, '/notificacion',
                      persona.pk, 1, 'sga', Persona)
jefrersion()
print("FIN ")


