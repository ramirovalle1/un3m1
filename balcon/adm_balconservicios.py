# -*- coding: UTF-8 -*-
import random
import sys
import json

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Q
import xlwt
from django.forms import model_to_dict
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.template import Context
from django.template.loader import get_template
from django.views.decorators.csrf import csrf_exempt
from xlwt import *
from django.shortcuts import render
from decorators import secure_module, last_access
from poli.forms import EncuestaPreguntaForm
from sagest.forms import DepartamentoForm, IntegranteDepartamentoForm, ResponsableDepartamentoForm, \
    SeccionDepartamentoForm
from sagest.funciones import encrypt_id, encuesta_objeto, crear_editar_encuesta
from sagest.models import Departamento, SeccionDepartamento, OpcionSistema
from settings import EMAIL_DOMAIN
from sga.commonviews import adduserdata
from .forms import ProcesoForm, CategoriaBalconForm, TipoBalconForm, RequisitoBalconForm, ServicioBalconForm, \
    AgenteForm, InformacionBalconForm, AgenteEditForm, RequisitoServicioForm, ResponsableForm, \
    ConfiguraServicioBalconForm, TipoProcesoServicioForm, EncuestaPreguntasBalconForm, EstrellasEncuestaForm
from sga.funciones import MiPaginador, log, generar_nombre
from sga.models import Administrativo, Persona
from .models import Proceso, Tipo, Categoria, Requisito, Servicio, Agente, Informacion, TIPO_INFORMACION, \
    ProcesoServicio, RequisitosConfiguracion, ResponsableDepartamento, ESTADO_SOLICITUD_BALCON, TipoProcesoServicio, \
    EncuestaProceso, PreguntaEncuestaProceso
from sga.templatetags.sga_extras import encrypt
from openpyxl import workbook as openxl
from openpyxl.styles import Font as openxlFont
from openpyxl.styles.alignment import Alignment as alin

@login_required(redirect_field_name='ret', login_url='/loginsagest')
@secure_module
@last_access
@transaction.atomic()
def view(request):
    data = {}
    adduserdata(request, data)
    usuario = request.user
    if request.method == 'POST':
        action = request.POST['action']

        if action == 'versolicitudes':
            try:
                data['ps'] = ps = ProcesoServicio.objects.get(pk=request.POST['id'])
                data['solicitudes'] = solicitudes = ps.solicitudes()
                estados = []
                for estado in ESTADO_SOLICITUD_BALCON:
                    estados.append({"id": estado[0],
                                    "nombre": estado[1],
                                    "total": solicitudes.filter(estado=estado[0]).count()})
                data['estados'] = estados
                data['tiene_novedades_admision'] = ps.tiene_novedades_admision()
                template = get_template("adm_balconservicios/estadistica/detalle.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'html': json_content})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": "Ocurrio un error al consultar los datos"})

        if action == 'loadSolicitudesByEstado':
            try:
                data['ps'] = ps = ProcesoServicio.objects.get(pk=request.POST['ps_id'])
                data['estado'] = estado = request.POST['estado']
                data['solicitudes'] = solicitudes = ps.solicitudes_por_estado(estado)
                template = get_template("adm_balconservicios/estadistica/solicitudes.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'html': json_content})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": "Ocurrio un error al consultar los datos"})

        if action == 'loadGraphicNoveltiesAdmisionByEstado':
            try:
                data['ps'] = ps = ProcesoServicio.objects.get(pk=request.POST['ps_id'])
                data['estado'] = estado = request.POST['estado']
                if not ps.tiene_novedades_admision_por_estado(estado):
                    raise NameError(u"No tiene registro de noveades")
                registros = ps.novedades_admision_por_estado(estado)
                tests = registros.exclude(examen__isnull=False)
                novedades_test = []
                for test in tests.order_by('test__titulo').values_list('test__titulo', flat=True).distinct():
                    novedades_test.append({"nombre": test if test else "TEST UNIDAD 3",
                                           "total": tests.filter(test__titulo=test).count()})
                examenes = registros.exclude(examen__isnull=True)
                novedades_examen = []
                for examen in examenes.order_by('examen__nombre').values_list('examen__nombre', flat=True).distinct():
                    novedades_examen.append({"nombre": examen,
                                             "total": examenes.filter(examen__nombre=examen).count()})
                data['novedades_test'] = novedades_test
                data['novedades_examen'] = novedades_examen
                data['novedades'] = (list(novedades_test) + list(novedades_examen))
                template = get_template("adm_balconservicios/estadistica/novedades_admision.html")
                json_content = template.render(data)
                return JsonResponse({"result": "ok", 'html': json_content})
            except Exception as ex:
                return JsonResponse({"result": "bad", "mensaje": "Ocurrio un error al consultar los datos"})

        if action == 'mostrarproceso':
            try:
                evento = Proceso.objects.get(pk=request.POST['id'])
                evento.activo = True if request.POST['val'] == 'y' else False
                evento.save(request)
                log(u'Visualiza Proceso en balcon : %s (%s)' % (evento, evento.activo),
                    request, "edit")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad"})

        if action == 'mostraradmin':
            try:
                evento = Proceso.objects.get(pk=request.POST['id'])
                evento.activoadmin = True if request.POST['val'] == 'y' else False
                evento.save(request)
                log(u'Visualiza Proceso administrador en balcon : %s (%s)' % (evento, evento.activo),
                    request, "edit")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad"})

        if action == 'mostrartipo':
            try:
                tipo = TipoProcesoServicio.objects.get(pk=request.POST['id'])
                tipo.mostrar = True if request.POST['val'] == 'y' else False
                tipo.save(request)
                log(u'Visualiza tipo de servicio en balcon : %s (%s)' % (tipo, tipo.mostrar),
                    request, "edit")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad"})

        if action == 'bloqueopublicacion':
            try:
                filtro = Informacion.objects.get(pk=int(encrypt(request.POST['id'])))
                filtro.mostrar = True if request.POST['val'] == 'y' else False
                filtro.save(request)
                log(u'Visualiza o no en Información en balcon : %s (%s)' % (filtro, filtro.mostrar),
                    request, "edit")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad"})

        if action == 'addproceso':
            try:
                f = ProcesoForm(request.POST)
                if f.is_valid():
                    if Proceso.objects.filter(status=True, sigla=request.POST['sigla']).exists():
                        transaction.set_rollback(True)
                        return JsonResponse({"result": "bad",
                                             "mensaje": u"SIGLAS {} YA SE ENCUENTRAN REGISTRADAS".format(
                                                 request.POST['sigla'])})
                    proceso = Proceso(sigla=f.cleaned_data['sigla'],
                                      descripcion=f.cleaned_data['descripcion'],
                                      interno=f.cleaned_data['interno'],
                                      externo=f.cleaned_data['externo'],
                                      tipo=f.cleaned_data['tipo'],
                                      categoria=f.cleaned_data['categoria'],
                                      departamento=f.cleaned_data['departamento'],
                                      subesolicitud=f.cleaned_data['subesolicitud']
                                      )
                    proceso.save(request)
                    log(u'Adiciono nuevo Proceso: %s' % proceso, request, "add")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos. %s" % ex})

        if action == 'editproceso':
            try:
                f = ProcesoForm(request.POST)
                if f.is_valid():
                    proceso = Proceso.objects.get(pk=int(encrypt(request.POST['id'])))
                    proceso.sigla = f.cleaned_data['sigla']
                    proceso.descripcion = f.cleaned_data['descripcion']
                    proceso.interno = f.cleaned_data['interno']
                    proceso.externo = f.cleaned_data['externo']
                    proceso.tipo = f.cleaned_data['tipo']
                    proceso.categoria = f.cleaned_data['categoria']
                    proceso.subesolicitud = f.cleaned_data['subesolicitud']
                    proceso.departamento = f.cleaned_data['departamento']
                    proceso.save(request)

                    log(u'Edito Proceso: %s' % proceso, request, "edit")
                    return JsonResponse({"result": "ok"})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al guardar los datos. %s" % ex})

        if action == 'delproceso':
            try:
                categoria = Proceso.objects.get(pk=request.POST['id'])
                categoria.status = False
                categoria.save(request)
                log(u'Elimino Proceso Balcon: %s' % categoria, request, "del")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        if action == 'addinformacion':
            try:
                form = InformacionBalconForm(request.POST, request.FILES)
                id = None
                newfile = None
                if 'archivomostrar' in request.FILES:
                    newfile = request.FILES['archivomostrar']
                    if newfile:
                        if newfile.size > 4194304:
                            return JsonResponse({"result": True, "mensaje": u"Error, archivo mayor a 3 Mb."})
                        else:
                            newfilesd = newfile._name
                            ext = newfilesd[newfilesd.rfind("."):]
                            if ext in ['.pdf', '.jpg', '.jpeg', '.png', '.jpeg', '.peg']:
                                newfile._name = generar_nombre("archivomostrar_", newfile._name)
                            else:
                                return JsonResponse(
                                    {"result": True, "mensaje": u"Error, Solo archivo con extención. pdf, jpg, jpeg."})
                newfile1 = None
                if 'archivodescargar' in request.FILES:
                    newfile1 = request.FILES['archivodescargar']
                    if newfile1:
                        if newfile1.size > 4194304:
                            return JsonResponse({"result": True, "mensaje": u"Error, archivo mayor a 3 Mb."})
                        else:
                            newfilesd = newfile1._name
                            ext = newfilesd[newfilesd.rfind("."):]
                            if ext in ['.pdf', '.jpg', '.jpeg', '.png', '.jpeg', '.peg']:
                                newfile1._name = generar_nombre("archivodescargar_", newfile1._name)
                            else:
                                return JsonResponse(
                                    {"result": True, "mensaje": u"Error, Solo archivo con extención. pdf, jpg, jpeg."})
                if form.is_valid():
                    id = int(encrypt(request.POST['id']))
                    if not ProcesoServicio.objects.filter(pk=id, status=True).exists():
                        return JsonResponse({'result': True, 'mensaje': 'No existe proceso servicio'})
                    info = Informacion(tipo=form.cleaned_data['tipo'],
                                              descripcion=form.cleaned_data['descripcion'],
                                              informacion=form.cleaned_data['informacion'],
                                              mostrar=form.cleaned_data['mostrar'],
                                              servicio_id = id)
                    if 'archivomostrar' in request.FILES:
                        info.archivomostrar = newfile
                    if 'archivodescargar' in request.FILES:
                        info.archivodescargar = newfile1
                    info.save(request)
                    log(u'Adiciono nueva informacion en el balcon: %s' % info, request, "add")
                    return JsonResponse({"result": False}, safe=False)
                else:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, 'mensaje': 'Error al guardar los datos'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': f'Error al guardar los datos. {ex}'})

        if action == 'editinformacion':
            try:
                form = InformacionBalconForm(request.POST, request.FILES)
                newfile = None
                if 'archivomostrar' in request.FILES:
                    newfile = request.FILES['archivomostrar']
                    if newfile:
                        if newfile.size > 4194304:
                            return JsonResponse({"result": True, "mensaje": u"Error, archivo mayor a 3 Mb."})
                        else:
                            newfilesd = newfile._name
                            ext = newfilesd[newfilesd.rfind("."):]
                            if ext in ['.pdf', '.jpg', '.jpeg', '.png', '.jpeg', '.peg']:
                                newfile._name = generar_nombre("archivomostrar", newfile._name)
                            else:
                                return JsonResponse(
                                    {"result": True, "mensaje": u"Error, Solo archivo con extención. pdf, jpg, jpeg."})
                newfile1 = None
                if 'archivodescargar' in request.FILES:
                    newfile1 = request.FILES['archivodescargar']
                    if newfile1:
                        if newfile1.size > 4194304:
                            return JsonResponse({"result": True, "mensaje": u"Error, archivo mayor a 3 Mb."})
                        else:
                            newfilesd = newfile1._name
                            ext = newfilesd[newfilesd.rfind("."):]
                            if ext in ['.pdf', '.jpg', '.jpeg', '.png', '.jpeg', '.peg']:
                                newfile1._name = generar_nombre("archivodescargar", newfile1._name)
                            else:
                                return JsonResponse(
                                    {"result": True, "mensaje": u"Error, Solo archivo con extención. pdf, jpg, jpeg."})
                info = Informacion.objects.get(pk=int(encrypt(request.POST['id'])))
                if form.is_valid():
                    info.tipo = form.cleaned_data['tipo']
                    info.descripcion = form.cleaned_data['descripcion']
                    info.informacion = form.cleaned_data['informacion']
                    info.mostrar = form.cleaned_data['mostrar']
                    if 'archivomostrar' in request.FILES:
                        info.archivomostrar = newfile
                    if 'archivodescargar' in request.FILES:
                        info.archivodescargar = newfile1
                    info.save(request)
                    log(u'Edito informacion en el balcon: %s' % info, request, "add")
                    return JsonResponse({"result": False}, safe=False)
                else:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, 'mensaje': 'Error al guardar los datos'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': f'Error al guardar los datos. {ex}'})

        if action == 'delinformacion':
            try:
                categoria = Informacion.objects.get(pk=request.POST['id'])
                categoria.status = False
                categoria.save(request)
                log(u'Elimino Información Balcon: %s' % categoria, request, "del")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        if action == 'delreqservicio':
            try:
                categoria = RequisitosConfiguracion.objects.get(pk=request.POST['id'])
                categoria.status = False
                categoria.save(request)
                log(u'Elimino Requisito Balcon: %s' % categoria, request, "del")
                return JsonResponse({"result": "ok"})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": "bad", "mensaje": u"Error al eliminar los datos."})

        if action == 'addrequisitoservicio':
            try:
                with transaction.atomic():
                    form = RequisitoServicioForm(request.POST)
                    if form.is_valid():
                        filtro = ProcesoServicio.objects.get(pk=request.POST['id'])
                        categoria = RequisitosConfiguracion(servicio=filtro, requisito=form.cleaned_data['requisito'],
                                                            obligatorio=form.cleaned_data['obligatorio'],
                                                            # leyenda=form.cleaned_data['leyenda'],
                                                            activo=form.cleaned_data['activo'])
                        categoria.save(request)
                        log(u'Adiciono Requisito en el balcon: %s' % categoria, request, "add")
                        # return JsonResponse({"result": False,'to':'/nuevaurl'}, safe=False) SI DESEAS REDIRECCIONAR ADICIONARLE TO A LA RESPUESTA
                        return JsonResponse({"result": False}, safe=False)
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({"result": True, "mensaje": "Complete los datos requeridos."}, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        if action == 'addtiposervicio':
            try:
                with transaction.atomic():
                    form = TipoProcesoServicioForm(request.POST)
                    if form.is_valid():
                        servicio = ProcesoServicio.objects.get(pk=request.POST['id'])
                        tipo = TipoProcesoServicio(servicio=servicio,
                                                   nombre=form.cleaned_data['nombre'].upper().strip(),
                                                   departamento=form.cleaned_data['departamento'],
                                                   mostrar=form.cleaned_data['mostrar'])
                        tipo.save(request)
                        log(u'Adiciono Tipo en el balcon: %s' % tipo, request, "add")
                        # return JsonResponse({"result": False,'to':'/nuevaurl'}, safe=False) SI DESEAS REDIRECCIONAR ADICIONARLE TO A LA RESPUESTA
                        return JsonResponse({"result": False}, safe=False)
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({"result": True, "mensaje": "Complete los datos requeridos."}, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        if action == 'editrequisitoservicio':
            try:
                with transaction.atomic():
                    filtro = RequisitosConfiguracion.objects.get(pk=request.POST['id'])
                    f = RequisitoServicioForm(request.POST)
                    if f.is_valid():
                        filtro.requisito = f.cleaned_data['requisito']
                        # filtro.leyenda = f.cleaned_data['leyenda']
                        filtro.obligatorio = f.cleaned_data['obligatorio']
                        filtro.activo = f.cleaned_data['activo']
                        filtro.save(request)
                        log(u'Modificó Categoria en balcon: %s' % filtro, request, "edit")
                        return JsonResponse({"result": False}, safe=False)
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({"result": True, "mensaje": "Complete los datos requeridos."}, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        if action == 'edittiposervicio':
            try:
                with transaction.atomic():
                    tipo = TipoProcesoServicio.objects.get(pk=request.POST['id'])
                    f = TipoProcesoServicioForm(request.POST)
                    if f.is_valid():
                        tipo.nombre = f.cleaned_data['nombre'].upper().strip()
                        tipo.departamento = f.cleaned_data['departamento']
                        tipo.descripcion = f.cleaned_data['descripcion']
                        tipo.mostrar = f.cleaned_data['mostrar']
                        tipo.save(request)
                        log(u'Modificó Tipo servicio en balcon: %s' % tipo, request, "edit")
                        return JsonResponse({"result": False}, safe=False)
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({"result": True, "mensaje": "Complete los datos requeridos."}, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Intentelo más tarde."}, safe=False)

        if action == 'addcategoriamodal':
            try:
                with transaction.atomic():
                    form = CategoriaBalconForm(request.POST)
                    if form.is_valid():
                        categoria = Categoria(descripcion=form.cleaned_data['descripcion'].upper(),
                                              estado=form.cleaned_data['estado'])
                        categoria.save(request)
                        for coordinacion in form.cleaned_data['coordinaciones']:
                            categoria.coordinaciones.add(coordinacion)

                        log(u'Adiciono Categoria en el balcon: %s' % categoria, request, "add")
                        # return JsonResponse({"result": False,'to':'/nuevaurl'}, safe=False) SI DESEAS REDIRECCIONAR ADICIONARLE TO A LA RESPUESTA
                        return JsonResponse({"result": False}, safe=False)
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({"result": True, "mensaje": "Complete los datos requeridos."}, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': f'Error al guardar los datos. {ex}'})

        if action == 'editcategoriamodal':
            try:
                with transaction.atomic():
                    filtro = Categoria.objects.get(pk=int(encrypt(request.POST['id'])))
                    form = CategoriaBalconForm(request.POST)
                    if form.is_valid():
                        filtro.descripcion = form.cleaned_data['descripcion'].upper()
                        filtro.estado = form.cleaned_data['estado']
                        filtro.coordinaciones.clear()
                        for coordinacion in form.cleaned_data['coordinaciones']:
                            filtro.coordinaciones.add(coordinacion)
                        filtro.save(request)
                        log(u'Modificó Categoria en balcon: %s' % filtro, request, "edit")
                        return JsonResponse({"result": False}, safe=False)
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({"result": True, "mensaje": "Complete los datos requeridos."}, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': f'Error al guardar los datos. {ex}'})

        if action == 'delcat':
            try:
                categoria = Categoria.objects.get(pk=encrypt(request.POST['id']), status=True)
                categoria.status = False
                categoria.save(request)
                log(u'Elimino Categoria Balcon: %s' % categoria, request, "del")
                res_json = {"error": False}
            except Exception as ex:
                transaction.set_rollback(True)
                res_json = {'error': True, "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)

        if action == 'addtipomodal':
            try:
                with transaction.atomic():
                    form = TipoBalconForm(request.POST)
                    if form.is_valid():
                        tipo = Tipo(descripcion=form.cleaned_data['descripcion'].upper(),
                                    estado=form.cleaned_data['estado'])
                        tipo.save(request)
                        log(u'Adiciono Tipo en el balcon: %s' % tipo, request, "add")
                        return JsonResponse({"result": False}, safe=False)
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({"result": True, "mensaje": "Complete los datos requeridos."}, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': f'Error al guardar los datos. {ex}'})

        if action == 'edittipomodal':
            try:
                with transaction.atomic():
                    filtro = Tipo.objects.get(pk=int(encrypt(request.POST['id'])))
                    form = TipoBalconForm(request.POST)
                    if form.is_valid():
                        filtro.descripcion = form.cleaned_data['descripcion'].upper()
                        filtro.estado = form.cleaned_data['estado']
                        filtro.save(request)
                        log(u'Modificó Tipo en balcon: %s' % filtro, request, "edit")
                        return JsonResponse({"result": False}, safe=False)
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({"result": True, "mensaje": "Complete los datos requeridos."}, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': f'Error al guardar los datos. {ex}'})

        if action == 'deltip':
            try:
                tipo = Tipo.objects.get(pk=encrypt(request.POST['id']), status=True)
                tipo.status = False
                tipo.save(request)
                log(u'Elimino Tipo Balcon: %s' % tipo, request, "del")
                res_json = {"error": False}
            except Exception as ex:
                transaction.set_rollback(True)
                res_json = {'error': True, "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)

        if action == 'addconfiguraservicio':
            try:
                with transaction.atomic():
                    form = ConfiguraServicioBalconForm(request.POST)
                    if form.is_valid():
                        proceso = Proceso.objects.get(pk=int(encrypt(request.POST['id'])))
                        if not ProcesoServicio.objects.filter(proceso=proceso,
                                                              servicio=form.cleaned_data['servicio']).exists():
                            filtro = ProcesoServicio(proceso=proceso,
                                                       servicio=form.cleaned_data['servicio'],
                                                       tiempominimo=form.cleaned_data['minimo'],
                                                       minutos=form.cleaned_data['minutos'],
                                                       tiempomaximo=form.cleaned_data['maximo'],
                                                       opcsistema=form.cleaned_data['opcsistema'],
                                                       url=form.cleaned_data['url'])
                            filtro.save(request)
                            log(u'Adiciono Servicio en balcon: %s' % filtro, request, "add")
                            return JsonResponse({"result": False}, safe=False)
                        else:
                            return JsonResponse({"result": True, "mensaje": "El servicio ya existe."}, safe=False)

                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({"result": True, "mensaje": "Complete los datos requeridos."}, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': f'Error al guardar los datos. {ex}'}, safe=False)

        if action == 'editservicio':
            try:
                with transaction.atomic():
                    form = ConfiguraServicioBalconForm(request.POST)
                    if form.is_valid():
                        filtro = ProcesoServicio.objects.get(pk=int(encrypt(request.POST['id'])))
                        filtro.tiempominimo = form.cleaned_data['minimo']
                        filtro.tiempomaximo = form.cleaned_data['maximo']
                        filtro.minutos = form.cleaned_data['minutos']
                        filtro.opcsistema = form.cleaned_data['opcsistema']
                        filtro.url = form.cleaned_data['url']
                        filtro.save(request)
                        log(u'Editó Servicio en balcon: %s' % filtro, request, "add")
                        return JsonResponse({"result": False}, safe=False)
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({"result": True, "mensaje": "Complete los datos requeridos."}, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': f'Error al guardar los datos. {ex}'})

        if action == 'addrequisitomodal':
            try:
                with transaction.atomic():
                    form = RequisitoBalconForm(request.POST)
                    if form.is_valid():
                        requisito = Requisito(descripcion=form.cleaned_data['descripcion'].upper(),
                                              estado=form.cleaned_data['estado'])
                        requisito.save(request)
                        log(u'Adiciono Requisito en el balcon: %s' % requisito, request, "add")
                        return JsonResponse({"result": False}, safe=False)
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({"result": True, "mensaje": "Complete los datos requeridos."}, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': f'Error al guardar los datos. {ex}'})

        if action == 'editrequisitomodal':
            try:
                with transaction.atomic():
                    filtro = Requisito.objects.get(pk=int(encrypt(request.POST['id'])))
                    form = RequisitoBalconForm(request.POST)
                    if form.is_valid():
                        filtro.descripcion = form.cleaned_data['descripcion'].upper()
                        filtro.estado = form.cleaned_data['estado']
                        filtro.save(request)
                        log(u'Modificó Requisito en balcon: %s' % filtro, request, "edit")
                        return JsonResponse({"result": False}, safe=False)
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({"result": True, "mensaje": "Complete los datos requeridos."}, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': f'Error al guardar los datos. {ex}'})

        if action == 'delreq':
            try:
                requisito = Requisito.objects.get(pk=encrypt(request.POST['id']), status=True)
                requisito.status = False
                requisito.save(request)
                log(u'Elimino Requisito Balcon: %s' % requisito, request, "del")
                res_json = {"error": False}
            except Exception as ex:
                transaction.set_rollback(True)
                res_json = {'error': True, "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)

        if action == 'addserviciomodal':
            try:
                with transaction.atomic():
                    form = ServicioBalconForm(request.POST)
                    if form.is_valid():
                        serviciob = Servicio(nombre=form.cleaned_data['nombre'].upper(),
                                             descripcion=form.cleaned_data['descripcion'].upper(),
                                             estado=form.cleaned_data['estado'])
                        lista = request.POST.getlist('opcsistema')
                        serviciob.save(request)
                        for l in lista:
                            opcsistema = OpcionSistema.objects.get(pk=int(l))
                            serviciob.opcsistema.add(opcsistema)
                        log(u'Adiciono Servicio en el balcon: %s' % serviciob, request, "add")
                        return JsonResponse({"result": False}, safe=False)
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({"result": True, "mensaje": "Complete los datos requeridos."}, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': f'Error al guardar los datos. {ex}'})

        if action == 'editserviciomodal':
            try:
                with transaction.atomic():
                    filtro = Servicio.objects.get(pk=int(encrypt(request.POST['id'])))
                    form = ServicioBalconForm(request.POST)
                    if form.is_valid():
                        filtro.nombre = form.cleaned_data['nombre'].upper()
                        filtro.descripcion = form.cleaned_data['descripcion'].upper()
                        filtro.estado = form.cleaned_data['estado']
                        opcsistemas = request.POST.getlist('opcsistema')
                        filtro.opcsistema.clear()
                        for l in opcsistemas:
                            opcsistema = OpcionSistema.objects.get(pk=int(l))
                            filtro.opcsistema.add(opcsistema)
                        filtro.save(request)
                        log(u'Modificó Servicio en balcon: %s' % filtro, request, "edit")
                        return JsonResponse({"result": False}, safe=False)
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({"result": True, "mensaje": "Complete los datos requeridos."}, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': f'Error al guardar los datos. {ex}'})

        if action == 'delser':
            try:
                serviciob = Servicio.objects.get(pk=encrypt(request.POST['id']), status=True)
                serviciob.status = False
                serviciob.save(request)
                log(u'Elimino Servicio Balcon: %s' % serviciob, request, "del")
                res_json = {"error": False}
            except Exception as ex:
                transaction.set_rollback(True)
                res_json = {'error': True, "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)

        if action == 'delservicio':
            try:
                fitlro = ProcesoServicio.objects.get(pk=int(encrypt(request.POST['id'])))
                fitlro.status = False
                fitlro.save(request)
                log(u'Elimino Servicio en Balcon: %s' % fitlro, request, "del")
                res_json = {"error": False}
            except Exception as ex:
                transaction.set_rollback(True)
                res_json = {'error': True, "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)


        if action == 'addpersonalmodal':
            try:
                with transaction.atomic():
                    if Agente.objects.filter(persona_id=int(request.POST['persona']), status=True).exists():
                        transaction.set_rollback(True)
                        return JsonResponse({"result": True, "mensaje": "Está persona ya se encuentra registrado."},
                                            safe=False)
                    form = AgenteForm(request.POST)
                    if form.is_valid():
                        agente = Agente(persona_id=int(request.POST['persona']),
                                        estado=form.cleaned_data['estado'],
                                        admin=form.cleaned_data['admin'])
                        agente.save(request)
                        log(u'Adiciono Agente en el balcon: %s' % agente, request, "add")
                        return JsonResponse({"result": False}, safe=False)
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({"result": True, "mensaje": "Complete los datos requeridos."}, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': f'Error al guardar los datos. {ex}'})

        if action == 'editpersonalmodal':
            try:
                with transaction.atomic():
                    filtro = Agente.objects.get(pk=int(encrypt(request.POST['id'])))
                    form = AgenteEditForm(request.POST)
                    if form.is_valid():
                        filtro.estado = form.cleaned_data['estado']
                        filtro.admin = form.cleaned_data['admin']
                        filtro.save(request)
                        log(u'Modificó Agente en balcon: %s' % filtro, request, "edit")
                        return JsonResponse({"result": False}, safe=False)
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({"result": True, "mensaje": "Complete los datos requeridos."}, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': f'Error al guardar los datos. {ex}'})

        if action == 'editresponsablemodal':
            try:
                with transaction.atomic():
                    filtro = ResponsableDepartamento.objects.get(pk=int(encrypt(request.POST['id'])))
                    form = ResponsableForm(request.POST)
                    if form.is_valid():
                        filtro.responsable_id = form.cleaned_data['responsable']
                        filtro.departamento = form.cleaned_data['departamento']
                        filtro.estado = form.cleaned_data['estado']
                        filtro.save(request)
                        log(u'Modificó responsable departamento en balcón de servicios: %s' % filtro, request, "edit")
                        return JsonResponse({"result": False}, safe=False)
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({"result": True, "mensaje": "Complete los datos requeridos."}, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': f'Error al guardar los datos. {ex}'})

        if action == 'delpersona':
            try:
                agente = Agente.objects.get(pk=encrypt(request.POST['id']), status=True)
                agente.status = False
                agente.save(request)
                log(u'Elimino Agente en Balcon: %s' % agente, request, "del")
                res_json = {"error": False}
            except Exception as ex:
                transaction.set_rollback(True)
                res_json = {'error': True, "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)

        if action == 'addresponsablemodal':
            try:
                with transaction.atomic():
                    if ResponsableDepartamento.objects.filter(responsable_id=int(request.POST['responsable']),
                                                              status=True).exists():
                        transaction.set_rollback(True)
                        return JsonResponse({"result": True, "mensaje": "Esta persona ya se encuentra registrado."},
                                            safe=False)
                    form = ResponsableForm(request.POST)
                    if form.is_valid():
                        responsable = ResponsableDepartamento(responsable_id=int(request.POST['responsable']),
                                                              departamento=form.cleaned_data['departamento'],
                                                              estado=form.cleaned_data['estado'])
                        responsable.save(request)
                        log(u'Adiciono responsable de balcon: %s' % responsable, request, "add")
                        return JsonResponse({"result": False}, safe=False)
                    else:
                        transaction.set_rollback(True)
                        return JsonResponse({"result": True, "mensaje": "Complete los datos requeridos."}, safe=False)
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': f'Error al guardar los datos. {ex}'})

        if action == 'delresponsable':
            try:
                filtro = ResponsableDepartamento.objects.get(pk=int(encrypt(request.POST['id'])))
                filtro.status = False
                filtro.save(request)
                log(u'Elimino Responsable en Balcon: %s' % filtro, request, "del")
                res_json = {"error": False}
            except Exception as ex:
                transaction.set_rollback(True)
                res_json = {'error': True, "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)

        if action == 'delpregunta':
            try:
                filtro = PreguntaEncuestaProceso.objects.get(pk=encrypt(request.POST['id']))
                filtro.status = False
                filtro.save(request)
                log(u'Elimino Pregunta de Encuesta en Balcon: %s' % filtro, request, "del")
                res_json = {"error": False}
            except Exception as ex:
                transaction.set_rollback(True)
                res_json = {'error': True, "message": "Error: {}".format(ex)}
            return JsonResponse(res_json, safe=False)

        if action == 'addpregunta':
            try:
                form = EncuestaPreguntasBalconForm(request.POST)
                encuesta = EncuestaProceso.objects.get(id=int(encrypt(request.POST['id'])))
                if form.is_valid():
                    if encuesta.preguntaencuestaproceso_set.filter(status=True,
                                                                   descripcion=form.cleaned_data[
                                                                       'descripcion']).exists():
                        raise NameError('Ya existe esta pregunta en esta encuesta')
                    filtro = PreguntaEncuestaProceso(encuesta=encuesta, descripcion=form.cleaned_data['descripcion'],
                                                     estado=form.cleaned_data['estado'])
                    filtro.save(request)
                    log(u'Registro pregunta balcón servicio: %s' % filtro, request, "add")
                    return JsonResponse({'result': False, 'mensaje': 'Registro Exitoso'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': f'Error al guardar los datos. {ex}'})

        if action == 'editpregunta':
            try:
                form = EncuestaPreguntasBalconForm(request.POST)
                filtro = PreguntaEncuestaProceso.objects.get(id=int(encrypt(request.POST['id'])))
                if form.is_valid():
                    if filtro.encuesta.preguntaencuestaproceso_set.filter(status=True, descripcion=form.cleaned_data[
                        'descripcion']).exclude(id=int(encrypt(request.POST['id']))).exists():
                        raise NameError('Ya existe esta pregunta en esta encuesta')
                    filtro.descripcion = form.cleaned_data['descripcion']
                    filtro.estado = form.cleaned_data['estado']
                    filtro.save(request)
                    log(u'Edito encuesta de balcón servicios: %s' % filtro, request, action)
                    return JsonResponse({'result': False, 'mensaje': 'Edicion Exitosa'})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': f'Error al guardar los datos. {ex}'})

        if action == 'estrellasencuesta':
            try:
                form = EstrellasEncuestaForm(request.POST)
                if form.is_valid():
                    # filtro = EncuestaProceso.objects.get(pk=int(encrypt(request.POST['id'])))
                    filtro = Proceso.objects.get(pk=int(encrypt(request.POST['id']))).encuestaproceso_set.filter(
                        status=True, vigente=True).last()
                    if int(form.cleaned_data['cantidad']):
                        if int(form.cleaned_data['cantidad']) > 10 or int(form.cleaned_data['cantidad']) <= 0:
                            raise NameError('Debe ingresar un número entre 0 o 5')
                    else:
                        raise NameError('Debe ingresar un número')
                    filtro.valoracion = form.cleaned_data['cantidad']
                    filtro.save(request)
                    log(u'Edito encuesta de balcón servicios: %s' % filtro, request, action)
                    return JsonResponse({'result': False, 'mensaje': 'Edicion Exitosa'})
                else:
                    raise NameError('Error')
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({'result': True, 'mensaje': f'Error al guardar los datos. {ex}'})

        elif action == 'addencuesta':
            with transaction.atomic():
                try:
                    id = encrypt_id(request.POST['id'])
                    form = EncuestaPreguntaForm(request.POST)
                    if not form.is_valid():
                        return JsonResponse({'result': True, "form": [{k: v[0]} for k, v in form.errors.items()], "message": "Error en el formulario"})

                    proceso = Proceso.objects.get(id=id)
                    encuesta = crear_editar_encuesta(request, proceso, form, 2)
                    preguntas = json.loads(request.POST['lista_items1'])
                    lista = []
                    for p in preguntas:
                        idpregunta = p['id_pregunta']
                        if not idpregunta:
                            pregunta = PreguntaEncuestaProceso(encuesta=encuesta,
                                                                estado=p['activo'],
                                                                descripcion=p['pregunta'])
                            pregunta.save(request)
                            log(u'Agrego pregunta : %s' % pregunta, request, "add")
                        else:
                            pregunta = PreguntaEncuestaProceso.objects.get(id=idpregunta)
                            if not pregunta.en_uso():
                                pregunta.descripcion = p['pregunta']
                            pregunta.estado = p['activo']
                            pregunta.save(request)
                            log(u'Edito pregunta : %s' % pregunta, request, "edit")
                        lista.append(pregunta.id)
                    PreguntaEncuestaProceso.objects.filter(status=True, encuesta=encuesta).exclude(id__in=lista).update(status=False)
                    return JsonResponse({'result': False, 'mensaje': u'Guardado con éxito'})
                except Exception as ex:
                    transaction.set_rollback(True)
                    return JsonResponse({'result': True, "mensaje": 'Error: {}'.format(ex)}, safe=False)

        return JsonResponse({"result": "bad", "mensaje": u"Solicitud Incorrecta."})
    else:
        if 'action' in request.GET:
            data['action'] = action = request.GET['action']

            if action == 'addproceso':
                try:
                    data['title'] = u'Adicionar proceso'
                    data['form'] = ProcesoForm()
                    data['requisitos'] = Requisito.objects.filter(status=True, estado=True)
                    return render(request, "adm_balconservicios/addproceso.html", data)
                except Exception as ex:
                    pass

            if action == 'editproceso':
                try:
                    data['title'] = u'Editar proceso'
                    data['proceso'] = proceso = Proceso.objects.get(pk=int(encrypt(request.GET['id'])))
                    servicios = Servicio.objects.filter(procesoservicio__proceso=proceso, procesoservicio__status=True)

                    data['form'] = ProcesoForm(initial={'sigla': proceso.sigla,
                                                        'descripcion': proceso.descripcion,
                                                        'tiempoestimado': proceso.tiempoestimado,
                                                        'tipo': proceso.tipo,
                                                        'categoria': proceso.categoria,
                                                        'departamento': proceso.departamento,
                                                        'servicio': servicios,
                                                        'responsable': proceso.persona,
                                                        'subesolicitud': proceso.subesolicitud,
                                                        'interno': proceso.interno,
                                                        'externo': proceso.externo,
                                                        })
                    return render(request, "adm_balconservicios/editproceso.html", data)
                except Exception as ex:
                    pass

            if action == 'addinformacion':
                try:
                    data['form'] = InformacionBalconForm()
                    data['id'] = int(encrypt(request.GET['id']))
                    data['action'] = action
                    template = get_template("adm_balconservicios/modal/forminformacion.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, 'message': str(ex)})

            if action == 'editinformacion':
                try:
                    data['id'] = id = int(encrypt(request.GET['id']))
                    data['action'] = action
                    filtro = Informacion.objects.get(pk=id)
                    initial = model_to_dict(filtro)
                    form = InformacionBalconForm(initial=initial)
                    data['form'] = form
                    template = get_template("adm_balconservicios/modal/forminformacion.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, 'message': str(ex)})

            if action == 'addcategoriamodal':
                try:
                    data['title'] = u'Ingresar Categoria'
                    data['action'] = request.GET['action']
                    form = CategoriaBalconForm()
                    data['form'] = form
                    template = get_template("adm_balconservicios/modal/formcategoria.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, 'message': str(ex)})

            if action == 'editcategoriamodal':
                try:
                    data['title'] = u'Editar Categoria'
                    data['id'] = id = int(encrypt(request.GET['id']))
                    data['action'] = request.GET['action']
                    data['filtro'] = filtro = Categoria.objects.get(id=id)
                    initial = model_to_dict(filtro)
                    form = CategoriaBalconForm(initial=initial)
                    data['form'] = form
                    template = get_template("adm_balconservicios/modal/formcategoria.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})

                except Exception as ex:
                    return JsonResponse({"result": False, 'message': str(ex)})

            if action == 'addtipomodal':
                try:
                    data['title'] = u'Ingresar Tipo'
                    data['action'] = request.GET['action']
                    form = TipoBalconForm()
                    data['form'] = form
                    template = get_template("adm_balconservicios/modal/formtipo.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, 'message': str(ex)})

            if action == 'edittipomodal':
                try:
                    data['title'] = u'Editar Tipo'
                    data['id'] = id = int(encrypt(request.GET['id']))
                    data['action'] = request.GET['action']
                    data['filtro'] = filtro = Tipo.objects.get(id=id)
                    initial = model_to_dict(filtro)
                    form = TipoBalconForm(initial=initial)
                    data['form'] = form
                    template = get_template("adm_balconservicios/modal/formtipo.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})

                except Exception as ex:
                    return JsonResponse({"result": False, 'message': str(ex)})

            if action == 'addconfiguraservicio':
                try:
                    data['form'] = ConfiguraServicioBalconForm()
                    data['id'] = int(encrypt(request.GET['id']))
                    template = get_template("adm_balconservicios/modal/formconfigservicio.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            if action == 'editservicio':
                try:
                    data['id'] = id = int(encrypt(request.GET['id']))
                    data['action'] = action
                    data['filtro'] = filtro = ProcesoServicio.objects.get(pk=id)
                    initial = model_to_dict(filtro)
                    form = ConfiguraServicioBalconForm(initial=initial)
                    data['form'] = form
                    template = get_template("adm_balconservicios/modal/formconfigservicio.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})

                except Exception as ex:
                    return JsonResponse({"result": False, 'message': str(ex)})

            if action == 'addrequisitomodal':
                try:
                    data['title'] = u'Ingresar Requisito'
                    data['action'] = request.GET['action']
                    form = RequisitoBalconForm()
                    data['form'] = form
                    template = get_template("adm_balconservicios/modal/formrequisito.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, 'message': str(ex)})

            if action == 'editrequisitomodal':
                try:
                    data['title'] = u'Editar Requisito'
                    data['id'] = id = int(encrypt(request.GET['id']))
                    data['action'] = request.GET['action']
                    data['filtro'] = filtro = Requisito.objects.get(id=id)
                    initial = model_to_dict(filtro)
                    form = RequisitoBalconForm(initial=initial)
                    data['form'] = form
                    template = get_template("adm_balconservicios/modal/formrequisito.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})

                except Exception as ex:
                    return JsonResponse({"result": False, 'message': str(ex)})

            if action == 'addserviciomodal':
                try:
                    data['title'] = u'Ingresar Servicio'
                    data['action'] = request.GET['action']
                    form = ServicioBalconForm()
                    data['form'] = form
                    template = get_template("adm_balconservicios/modal/formservicio.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, 'message': str(ex)})

            if action == 'editserviciomodal':
                try:
                    data['title'] = u'Editar Servicio'
                    data['id'] = id = int(encrypt(request.GET['id']))
                    data['action'] = request.GET['action']
                    data['filtro'] = filtro = Servicio.objects.get(id=id)
                    initial = model_to_dict(filtro)
                    form = ServicioBalconForm(initial=initial)
                    data['form'] = form
                    template = get_template("adm_balconservicios/modal/formservicio.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})

                except Exception as ex:
                    return JsonResponse({"result": False, 'message': str(ex)})

            if action == 'addpersonalmodal':
                try:
                    data['title'] = u'Ingresar Agente'
                    data['action'] = request.GET['action']
                    form = AgenteForm()
                    form.fields['persona'].queryset = Persona.objects.none()
                    data['form'] = form
                    template = get_template("adm_balconservicios/modal/formagente.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, 'message': str(ex)})

            if action == 'editpersonalmodal':
                try:
                    data['title'] = u'Editar Agente'
                    data['id'] = id = int(encrypt(request.GET['id']))
                    data['action'] = request.GET['action']
                    data['filtro'] = filtro = Agente.objects.get(id=id)
                    initial = model_to_dict(filtro)
                    form = AgenteEditForm(initial=initial)
                    data['form'] = form
                    template = get_template("adm_balconservicios/modal/formagente.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})

                except Exception as ex:
                    return JsonResponse({"result": False, 'message': str(ex)})

            if action == 'editresponsablemodal':
                try:
                    data['title'] = u'Editar Responsable'
                    data['id'] = id = int(encrypt(request.GET['id']))
                    data['action'] = request.GET['action']
                    data['filtro'] = filtro = ResponsableDepartamento.objects.get(id=id)
                    initial = model_to_dict(filtro)
                    form = ResponsableForm(initial=initial)
                    form.fields['responsable'].queryset = Persona.objects.filter(id=filtro.responsable_id)
                    data['form'] = form
                    template = get_template("adm_balconservicios/modal/formresponsable.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, 'message': str(ex)})

            if action == 'addresponsablemodal':
                try:
                    data['title'] = u'Ingresar Responsable'
                    data['action'] = request.GET['action']
                    form = ResponsableForm()
                    form.fields['responsable'].queryset = Persona.objects.none()
                    data['form'] = form
                    template = get_template("adm_balconservicios/modal/formresponsable.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, 'message': str(ex)})

            if action == 'vistaprevia':
                try:
                    data['id'] = request.GET['id']
                    data['filtro'] = Informacion.objects.get(pk=int(encrypt(request.GET['id'])))
                    template = get_template("adm_balconservicios/modal/vistaprevia.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, 'message': str(ex)})

            if action == 'verdetalle':
                try:
                    data['id'] = request.GET['id']
                    data['filtro'] = proceso = Proceso.objects.get(pk=request.GET['id'])
                    data['servicios'] = ProcesoServicio.objects.filter(proceso=proceso, status=True)

                    template = get_template("adm_balconservicios/modal/detalleproceso.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            if action == 'viewinformacion':
                try:
                    url_vars = ''
                    id = int(encrypt(request.GET['id']))
                    if not ProcesoServicio.objects.filter(status = True, pk= id).exists():
                        raise NameError(u"Configuración no encontrada")
                    servicio = ProcesoServicio.objects.get(status=True, pk=id)
                    filtro = Q(status=True) & Q (servicio = servicio)
                    data['title'] = u'Informacion'
                    search = None
                    ids = None
                    tipo = None

                    if 't' in request.GET:
                        if request.GET['t'] != '0':
                            tipo = request.GET['t']
                            if tipo:
                                filtro = filtro & Q(tipo=int(tipo))
                                url_vars += '&t=' + tipo
                    if 's' in request.GET:
                        if request.GET['s'] != '':
                            search = request.GET['s'].strip().upper()
                            if search:
                                filtro = filtro & (Q(descripcion__icontains=search))
                                url_vars += '&s=' + search

                    procesos = Informacion.objects.filter(filtro).order_by('-id')
                    paging = MiPaginador(procesos, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['ids'] = ids if ids else ""
                    data['proceso'] = page.object_list
                    data['email_domain'] = EMAIL_DOMAIN
                    url_vars += '&action={}'.format(action)+'&id={}'.format(request.GET['id'])
                    data['search'] = search if search else ""
                    data['t'] = int(tipo) if tipo else ''
                    data["url_vars"] = url_vars
                    data["tipoinfo"] = TIPO_INFORMACION
                    data['servicio'] = servicio
                    return render(request, "adm_balconservicios/viewinformacion.html", data)
                except Exception as ex:
                    print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno))
                    pass

            elif action == 'buscarpersona':
                try:
                    q = request.GET['q'].upper().strip()
                    s = q.split(" ")
                    dep = Departamento.objects.get(status=True, pk=int(request.GET['gdep']))
                    if len(s) == 1:
                        per = dep.integrantes.filter((Q(nombres__icontains=q) | Q(apellido1__icontains=q) | Q(
                            apellido2__icontains=q) | Q(cedula__contains=q)),
                                                     Q(status=True)).distinct()[:15]
                    elif len(s) == 2:
                        per = dep.integrantes.filter((Q(apellido1__contains=s[0]) & Q(apellido2__contains=s[1])) | (
                                Q(nombres__icontains=s[0]) & Q(
                            nombres__icontains=s[1])) | (
                                                             Q(nombres__icontains=s[0]) & Q(
                                                         apellido1__contains=s[1]))).filter(status=True).distinct()[
                              :15]
                    else:
                        per = dep.integrantes.filter((Q(nombres__contains=s[0]) & Q(apellido1__contains=s[1]) & Q(
                            apellido2__contains=s[2])) | (Q(nombres__contains=s[0]) & Q(
                            nombres__contains=s[1]) & Q(apellido1__contains=s[2]))).filter(
                            status=True).distinct()[:15]

                    data = {"result": "ok",
                            "results": [{"id": x.id, "name": str(x.nombre_completo())}
                                        for x in per]}
                    return JsonResponse(data)
                except Exception as ex:
                    pass

            elif action == 'buscarservicio':
                try:
                    q = request.GET['q'].upper().strip()
                    s = q.split(" ")
                    pro = Proceso.objects.get(status=True, pk=int(request.GET['gdep']))
                    servicios = ProcesoServicio.objects.filter(proceso=pro, status=True)
                    data = {"result": "ok",
                            "results": [{"id": x.id, "name": str(x)}
                                        for x in servicios]}
                    return JsonResponse(data)
                except Exception as ex:
                    pass

            elif action == 'buscarpersona2':
                try:
                    q = request.GET['q'].upper().strip()
                    s = q.split(" ")
                    if len(s) == 1:
                        per = Persona.objects.filter(
                            (Q(nombres__icontains=q) | Q(apellido1__icontains=q) | Q(cedula__icontains=q) | Q(
                                apellido2__icontains=q) | Q(cedula__contains=q)),
                            Q(status=True)).distinct()[:15]
                    elif len(s) == 2:
                        per = Persona.objects.filter(
                            (Q(apellido1__contains=s[0]) & Q(apellido2__contains=s[1])) | (
                                    Q(nombres__icontains=s[0]) & Q(
                                nombres__icontains=s[1])) | (
                                    Q(nombres__icontains=s[0]) & Q(
                                apellido1__contains=s[1]))).filter(status=True).distinct()[
                              :15]
                    else:
                        per = Persona.objects.filter(
                            (Q(nombres__contains=s[0]) & Q(apellido1__contains=s[1]) & Q(
                                apellido2__contains=s[2])) | (Q(nombres__contains=s[0]) & Q(
                                nombres__contains=s[1]) & Q(apellido1__contains=s[2]))).filter(
                            status=True).distinct()[:15]

                    data = {"result": "ok",
                            "results": [{"id": x.id, "name": "{} - {}".format(x.cedula, x.nombre_completo())}
                                        for x in per]}
                    return JsonResponse(data)
                except Exception as ex:
                    pass

            if action == 'delinformacion':
                try:
                    data['title'] = u'ELIMINAR INFORMACIÓN'
                    data['filtro'] = Informacion.objects.get(pk=request.GET['id'])
                    return render(request, 'adm_balconservicios/delinformacion.html', data)
                except Exception as ex:
                    pass

            if action == 'delproceso':
                try:
                    data['title'] = u'ELIMINAR PROCESO'
                    data['filtro'] = Proceso.objects.get(pk=request.GET['id'])
                    return render(request, 'adm_balconservicios/delproceso.html', data)
                except Exception as ex:
                    pass

            if action == 'viewcategoria':
                try:
                    data['title'] = u'Categoria'
                    search = None
                    ids = None
                    url_vars = f"&action=viewcategoria"
                    if 's' in request.GET:
                        search = request.GET['s'].upper()
                        categoria = Categoria.objects.filter(descripcion__icontains=search, status=True).order_by('pk')
                        url_vars += f"&s={search}"
                    else:
                        categoria = Categoria.objects.filter(status=True).order_by('pk')
                    paging = MiPaginador(categoria, 20)

                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    request.session['viewactivobalconservicios'] = 1
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['categoria'] = page.object_list
                    data['url_vars'] = url_vars
                    return render(request, 'adm_balconservicios/viewcategoria.html', data)
                except Exception as ex:
                    pass

            if action == 'viewtipo':
                try:
                    data['title'] = u'Tipo'
                    search = None
                    ids = None
                    url_vars = f"&action=viewtipo"
                    if 's' in request.GET:
                        search = request.GET['s'].upper()
                        tipo = Tipo.objects.filter(descripcion__icontains=search, status=True).order_by('pk')
                        url_vars += f"&s={search}"
                    else:
                        tipo = Tipo.objects.filter(status=True).order_by('pk')
                    paging = MiPaginador(tipo, 20)

                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    request.session['viewactivobalconservicios'] = 2
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['tipo'] = page.object_list
                    data['url_vars'] = url_vars
                    return render(request, 'adm_balconservicios/viewtipo.html', data)
                except Exception as ex:
                    pass

            if action == 'viewrequisito':
                try:
                    data['title'] = u'Requisito'
                    search = None
                    ids = None
                    url_vars = f"&action=viewrequisito"
                    if 's' in request.GET:
                        search = request.GET['s'].upper()
                        requisito = Requisito.objects.filter(descripcion__icontains=search, status=True).order_by('pk')
                        url_vars += f"&s={search}"
                    else:
                        requisito = Requisito.objects.filter(status=True).order_by('pk')
                    paging = MiPaginador(requisito, 20)

                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    request.session['viewactivobalconservicios'] = 3
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['requisito'] = page.object_list
                    data['url_vars'] = url_vars
                    return render(request, 'adm_balconservicios/viewrequisito.html', data)
                except Exception as ex:
                    pass

            if action == 'viewservicio':
                try:
                    data['title'] = u'Servicio'
                    search = None
                    ids = None
                    url_vars = f"&action=viewservicio"
                    if 's' in request.GET:
                        search = request.GET['s'].upper()
                        servicio = Servicio.objects.filter(
                            Q(nombre__icontains=search) | Q(descripcion__icontains=search), status=True).order_by('pk')
                        url_vars += f"&s={search}"
                    else:
                        servicio = Servicio.objects.filter(status=True).order_by('pk')
                    paging = MiPaginador(servicio, 20)

                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    request.session['viewactivobalconservicios'] = 4
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['servicios'] = page.object_list
                    data['url_vars'] = url_vars
                    return render(request, 'adm_balconservicios/viewservicio.html', data)
                except Exception as ex:
                    pass

            if action == 'viewagentes':
                try:
                    data['title'] = u'Agentes'
                    search = None
                    ids = None
                    url_vars = f"&action=viewagentes"
                    if 's' in request.GET:
                        search = request.GET['s'].upper()
                        ss = search.split(' ')
                        if len(ss) == 1:
                            agente = Agente.objects.filter(
                                Q(persona__apellido1__icontains=search) | Q(persona__apellido2__icontains=search) | Q(
                                    persona__nombres__icontains=search), status=True).order_by('persona__apellido1')
                        else:
                            agente = Agente.objects.filter(
                                Q(persona__apellido1__icontains=ss[0]) | Q(persona__apellido2__icontains=ss[1]),
                                status=True).order_by('persona__apellido1')
                        url_vars += f"&s={search}"
                    else:
                        agente = Agente.objects.filter(status=True).order_by('pk')
                    paging = MiPaginador(agente, 20)

                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    request.session['viewactivobalconservicios'] = 5
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['personal'] = page.object_list
                    data['url_vars'] = url_vars
                    return render(request, 'adm_balconservicios/viewagentes.html', data)
                except Exception as ex:
                    pass

            if action == 'viewresponsables':
                try:
                    data['title'] = u'Responsables'
                    search = None
                    ids = None
                    url_vars = f"&action=viewresponsables"
                    if 's' in request.GET:
                        search = request.GET['s'].upper()
                        ss = search.split(' ')
                        if len(ss) == 1:
                            responsable = ResponsableDepartamento.objects.filter(
                                Q(responsable__apellido1__icontains=search) | Q(
                                    responsable__apellido2__icontains=search) | Q(
                                    responsable__nombres__icontains=search) | Q(departamento__nombre__icontains=search),
                                status=True).order_by('responsable__apellido1')
                        else:
                            responsable = ResponsableDepartamento.objects.filter(
                                Q(responsable__apellido1__icontains=ss[0]) | Q(responsable__apellido2__icontains=ss[1]),
                                status=True).order_by('responsable__apellido1')
                        url_vars += f"&s={search}"
                    else:
                        responsable = ResponsableDepartamento.objects.filter(status=True).order_by('pk')
                    paging = MiPaginador(responsable, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    request.session['viewactivobalconservicios'] = 6
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data['ids'] = ids if ids else ""
                    data['responsables'] = page.object_list
                    data['url_vars'] = url_vars
                    return render(request, 'adm_balconservicios/viewresponsables.html', data)
                except Exception as ex:
                    pass

            if action == 'configurarequisitos':
                try:
                    data['title'] = u'Configurar requisitos'
                    id = int(encrypt(request.GET['id']))
                    if not Proceso.objects.filter(pk=id).exists():
                        raise NameError(u"Proceso no encontrado")
                    url_vars = '&action=' + action + '&id=' + request.GET['id']
                    ids = None
                    search = None
                    proceso = Proceso.objects.get(pk=id)
                    filtro = Q(status=True, proceso=proceso)
                    if 's' in request.GET:
                        if request.GET['s'] != '':
                            search = request.GET['s'].strip().upper()
                            if search:
                                filtro = filtro & (Q(servicio__nombre__icontains=search))
                                url_vars += '&s=' + search
                    servicios = ProcesoServicio.objects.filter(filtro).order_by('id')
                    paging = MiPaginador(servicios, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data["url_vars"] = url_vars
                    data['ids'] = ids if ids else ""
                    data['id'] = request.GET['id']
                    data['servicios'] = page.object_list
                    data['proceso'] = proceso
                    return render(request, "adm_balconservicios/configurarequisitos.html", data)
                except Exception as ex:
                    return HttpResponseRedirect("/adm_balconservicios?info=%s" % ex)

            if action == 'configuratipo':
                try:
                    data['title'] = u'Configurar tipo'
                    data['servicio'] = servicio = ProcesoServicio.objects.get(pk=request.GET['id'])
                    data['tipos'] = TipoProcesoServicio.objects.filter(status=True, servicio=servicio).order_by('id')
                    return render(request, "adm_balconservicios/viewtiposervicio.html", data)
                except Exception as ex:
                    pass

            if action == 'requisitosservicio':
                try:
                    data['servicio'] = servicio = ProcesoServicio.objects.get(pk=request.GET['id'])
                    data['title'] = u'Requisitos Servicio'
                    data['listado'] = servicio.requisitosconfiguracion_set.filter(status=True).order_by('pk')
                    return render(request, "adm_balconservicios/requisitosservicio.html", data)
                except Exception as ex:
                    pass

            if action == 'addrequisitoservicio':
                try:
                    data['id'] = id = request.GET['id']
                    data['servicio'] = servicio = ProcesoServicio.objects.get(pk=request.GET['id'])
                    data['form2'] = RequisitoServicioForm()
                    template = get_template("adm_balconservicios/modal/formtipoprocesoservicio.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            if action == 'addtiposervicio':
                try:
                    data['id'] = id = request.GET['id']
                    data['servicio'] = servicio = ProcesoServicio.objects.get(pk=request.GET['id'])
                    data['form2'] = TipoProcesoServicioForm()
                    template = get_template("adm_balconservicios/modal/formrequisitoservicio.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            if action == 'editrequisitoservicio':
                try:
                    data['id'] = request.GET['id']
                    data['filtro'] = filtro = RequisitosConfiguracion.objects.get(pk=request.GET['id'])
                    data['form2'] = RequisitoServicioForm(initial=model_to_dict(filtro))
                    template = get_template("adm_balconservicios/modal/formrequisitoservicio.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            if action == 'edittiposervicio':
                try:
                    data['id'] = request.GET['id']
                    data['tipo'] = tipo = TipoProcesoServicio.objects.get(pk=request.GET['id'])
                    data['form2'] = TipoProcesoServicioForm(initial=model_to_dict(tipo))
                    template = get_template("adm_balconservicios/modal/formtipoprocesoservicio.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            if action == 'delreqservicio':
                try:
                    data['title'] = u'ELIMINAR REQUISITO'
                    data['requisito'] = RequisitosConfiguracion.objects.get(pk=request.GET['id'])
                    return render(request, 'adm_balconservicios/delrequisitoservicio.html', data)
                except Exception as ex:
                    pass

            if action == 'estadistica':
                try:
                    data['title'] = u'Estadisticas'
                    if not Proceso.objects.filter(pk=request.GET['id']).exists():
                        raise NameError(u"Proceso no encontrado")
                    data['proceso'] = proceso = Proceso.objects.get(pk=request.GET['id'])
                    return render(request, 'adm_balconservicios/estadistica/view.html', data)
                except Exception as ex:
                    return HttpResponseRedirect("/adm_balconservicios?info=%s" % ex)

            if action == 'encuestaproceso':
                try:
                    data['title'] = u'Encuesta de satisfacción'
                    search = None
                    url_vars = f"&action=encuestaproceso&id=" + request.GET['id']
                    id = int(encrypt(request.GET['id']))
                    if not Proceso.objects.filter(pk=id).exists():
                        raise NameError(u"Proceso no encontrado")
                    proceso = Proceso.objects.get(pk=id)
                    if not proceso.tiene_encuesta():
                        encuesta = EncuestaProceso(proceso=proceso, valoracion=5, vigente=True, categoria_id=2)
                        encuesta.save(request)
                    else:
                        encuesta = EncuestaProceso.objects.filter(proceso=proceso, status=True, vigente=True).last()
                    filtro = Q(status=True) & Q(encuesta=encuesta)
                    if 's' in request.GET:
                        search = request.GET['s'].strip().upper()
                        if search:
                            filtro = filtro & (
                                Q(descripcion__icontains=search))
                            url_vars += '&s=' + search
                    preguntas = PreguntaEncuestaProceso.objects.filter(filtro)
                    paging = MiPaginador(preguntas, 20)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                            p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['search'] = search if search else ""
                    data["url_vars"] = url_vars
                    data['preguntas'] = page.object_list
                    data['encuesta'] = encuesta
                    return render(request, 'adm_balconservicios/encuesta/view.html', data)
                except Exception as ex:
                    return HttpResponseRedirect("/adm_balconservicios?info=%s" % ex)

            if action == 'addpregunta':
                try:
                    data['title'] = u'Ingresar Preguntar'
                    data['action'] = request.GET['action']
                    data['id'] = int(encrypt(request.GET['id']))
                    form = EncuestaPreguntasBalconForm()
                    data['form'] = form
                    template = get_template("adm_balconservicios/encuesta/preguntas.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, 'mensaje': str(ex)})

            if action == 'editpregunta':
                try:
                    data['title'] = u'Editar Pregunta Encuesta de satisfacción'
                    data['id'] = id = int(encrypt(request.GET['id']))
                    data['action'] = action
                    filtro = PreguntaEncuestaProceso.objects.get(id=id)
                    initial = model_to_dict(filtro)
                    form = EncuestaPreguntasBalconForm(initial=initial)
                    data['form'] = form
                    template = get_template("adm_balconservicios/encuesta/preguntas.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})

                except Exception as ex:
                    return JsonResponse({"result": False, 'mensaje': str(ex)})

            if action == 'estrellasencuesta':
                try:
                    data['id'] = id = int(encrypt(request.GET['id']))
                    data['action'] = action
                    if not Proceso.objects.filter(pk=id).exists():
                        return JsonResponse({"result": False, 'mensaje': u"Proceso no encontrado"})
                    filtro = Proceso.objects.get(pk=id).encuestaproceso_set.filter(status=True, vigente=True)
                    if not filtro.exists():
                        return JsonResponse({"result": False, 'mensaje': u"No posee una encuesta vigente"})
                    filtro = filtro.last()
                    initial = model_to_dict(filtro)
                    form = EstrellasEncuestaForm(initial=initial)
                    form.fields['cantidad'].initial = filtro.valoracion
                    data['form'] = form
                    template = get_template("adm_balconservicios/encuesta/preguntas.html")
                    json_content = template.render(data)
                    return JsonResponse({"result": True, 'data': json_content})
                except Exception as ex:
                    return JsonResponse({"result": False, 'mensaje': str(ex)})

            elif action == 'addencuesta':
                try:
                    data['id'] = id = encrypt_id(request.GET['id'])
                    proceso = Proceso.objects.get(id=id)
                    encuesta = encuesta_objeto(proceso).first()
                    if encuesta:
                        form = EncuestaPreguntaForm(initial=model_to_dict(encuesta))
                        data['preguntas'] = PreguntaEncuestaProceso.objects.filter(status=True, encuesta=encuesta)
                    else:
                        form = EncuestaPreguntaForm()
                    data['form'] = form
                    data['switchery'] = True
                    template = get_template('formencuesta.html')
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    messages.error(request, f'{ex}')

            elif action == 'descargarreporte':
                try:
                    wb = openxl.Workbook()
                    wb["Sheet"].title = "Reporte_procesos"
                    ws = wb.active
                    style_title = openxlFont(name='Arial', size=16, bold=True)
                    style_cab = openxlFont(name='Arial', size=10, bold=True)
                    alinear = alin(horizontal="center", vertical="center")
                    response = HttpResponse(content_type="application/ms-excel")
                    response[
                        'Content-Disposition'] = 'attachment; filename=Reporte de procesos' + '-' + random.randint(
                        1, 10000).__str__() + '.xlsx'
                    ws.column_dimensions['B'].width = 20
                    ws.column_dimensions['C'].width = 25
                    ws.column_dimensions['D'].width = 20
                    ws.column_dimensions['E'].width = 30
                    ws.column_dimensions['F'].width = 25
                    ws.column_dimensions['G'].width = 20
                    ws.merge_cells('A1:M1')
                    ws['A1'] = 'REPORTE DE PROCESOS DE BALCÓN DE SERVICIOS'
                    celda1 = ws['A1']
                    celda1.font = style_title
                    celda1.alignment = alinear

                    columns = [u"N°", u"Estado publicado", "Sigla",
                               u"Descripción", u"Tipo", "Categoría", "Dirección",
                               "Encargado", "Tiempo Estimado", "¿Para Internos?",
                               "¿Para Externos?", "Fecha creación", "¿Encuesta configurada?"
                               ]
                    row_num = 3
                    for col_num in range(0, len(columns)):
                        celda = ws.cell(row=row_num, column=(col_num + 1), value=columns[col_num])
                        celda.font = style_cab

                    mensaje = 'NO REGISTRA'
                    row_num = 4
                    numero = 0
                    filtro = Q(status=True)
                    listado = Proceso.objects.filter(filtro).order_by('-fecha_creacion')
                    for l in listado:
                        numero += 1
                        estadopublicado = "Vigente" if l.informacion_mostrada().exists() else "No vigente"
                        interno = "Si" if l.interno else "No"
                        externo = "Si" if l.externo else "No"
                        encuestaconfigurada = "Si" if l.encuesta_configurada() else "No"
                        ws.cell(row=row_num, column=1, value=numero)
                        ws.cell(row=row_num, column=2, value=estadopublicado)
                        ws.cell(row=row_num, column=3, value=l.sigla)
                        ws.cell(row=row_num, column=4, value=l.descripcion)
                        ws.cell(row=row_num, column=5, value=str(l.tipo))
                        ws.cell(row=row_num, column=6, value=str(l.categoria))
                        ws.cell(row=row_num, column=7, value=str(l.departamento))
                        ws.cell(row=row_num, column=8, value=str(l.persona))
                        ws.cell(row=row_num, column=9, value=l.tiempoestimado)
                        ws.cell(row=row_num, column=10, value=interno)
                        ws.cell(row=row_num, column=11, value=externo)
                        ws.cell(row=row_num, column=12, value=str(l.fecha_creacion))
                        ws.cell(row=row_num, column=13, value=encuestaconfigurada)
                        row_num += 1
                    wb.save(response)
                    return response
                except Exception as ex:
                    messages.error(request, f'Error: {ex}')

            return HttpResponseRedirect(request.path)
        else:
            data['title'] = u'Procesos de la institución'

            url_vars = ''
            filtro = Q(status=True)
            search = None
            ids = None
            tipo = None

            if 't' in request.GET:
                if request.GET['t'] != '0':
                    tipo = request.GET['t']
            if 's' in request.GET:
                if request.GET['s'] != '':
                    search = request.GET['s'].upper()

            if search:
                filtro = filtro & (Q(descripcion__icontains=search) | (Q(departamento__nombre__icontains=search)))
                url_vars += '&s=' + search

            if tipo:
                filtro = filtro & Q(categoria_id=int(tipo))
                url_vars += '&t=' + tipo

            procesos = Proceso.objects.filter(filtro).order_by('id')

            paging = MiPaginador(procesos, 20)
            p = 1
            try:
                paginasesion = 1
                if 'paginador' in request.session:
                    paginasesion = int(request.session['paginador'])
                if 'page' in request.GET:
                    p = int(request.GET['page'])
                    p = int(request.GET['page'])
                else:
                    p = paginasesion
                try:
                    page = paging.page(p)
                except:
                    p = 1
                page = paging.page(p)
            except:
                page = paging.page(p)
            request.session['paginador'] = p
            data['paging'] = paging
            data['rangospaging'] = paging.rangos_paginado(p)
            data['page'] = page
            data['search'] = search if search else ""
            data['t'] = int(tipo) if tipo else ''
            data["url_vars"] = url_vars
            data['ids'] = ids if ids else ""
            data['proceso'] = page.object_list
            data['email_domain'] = EMAIL_DOMAIN
            data['categorias'] = Categoria.objects.filter(status=True, estado=True).order_by('descripcion')
            return render(request, 'adm_balconservicios/view.html', data)
