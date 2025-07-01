# coding=utf-8
# !/usr/bin/env python

import os
import sys

# SITE_ROOT = os.path.dirname(os.path.realpath(__file__))

YOUR_PATH = os.path.dirname(os.path.realpath(__file__))
# print(f"YOUR_PATH: {YOUR_PATH}")
SITE_ROOT = os.path.dirname(os.path.dirname(YOUR_PATH))
SITE_ROOT = os.path.join(SITE_ROOT, '')
# print(f"SITE_ROOT: {SITE_ROOT}")
your_djangoproject_home = os.path.split(SITE_ROOT)[0]
# print(f"your_djangoproject_home: {your_djangoproject_home}")
sys.path.append(your_djangoproject_home)

import xlwt
from webpush import send_user_notification
from xlwt import easyxf
from django.core.wsgi import get_wsgi_application

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "settings")
application = get_wsgi_application()
from datetime import datetime, timedelta
from django.db import transaction
from sga.models import Materia, Clase, Leccion, LeccionGrupo, AsistenciaLeccion, CamposTitulosPostulacion
from sga.funciones import variable_valor, notificacion
# from sga.models import *
# from sagest.models import *
import xlrd
from postulate.models import Partida, PersonaAplicarPartida, PersonaFormacionAcademicoPartida, PartidaTribunal
from sga.models import *
from sagest.models import *
from django.db.models import Sum, F, FloatField, IntegerField
from voto.models import ConfiguracionMesaResponsable
from django.db.models.functions import Coalesce
from settings import MEDIA_ROOT, BASE_DIR
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, colors
from django.http import HttpResponse
from sga.models import Modulo
from gdocumental.models import *
from bd.models import *
from moodle import moodle
from postulate.models import *
from core.firmar_documentos import verificarFirmasPDF
from inno.models import InformeMensualDocente


# def validacion_informemensual():
#     try:
#         informesinvalidos = []
#         qsbase_ = InformeMensualDocente.objects.filter(status=True, distributivo__periodo__id=177, fechafin__month=6, archivo__isnull=False)
#         num_ = qsbase_.count()
#         print(f"Leyendo {num_}")
#         inicia = 1
#         for informe in qsbase_:
#             valido, msg, datos = verificarFirmasPDF(informe.archivo)
#             if not valido:
#                 print(f"{inicia}/{num_} {informe.distributivo.profesor.persona.cedula} - {informe.distributivo.__str__()}")
#                 generado = 'informemensual_' + str(informe.distributivo.id) + '_' + str(informe.fechafin.month)
#                 firmado = 'informemensual_' + str(informe.distributivo.id) + '_' + str(informe.fechafin.month) + '_2'
#                 folder = os.path.join(SITE_STORAGE, 'media', 'informemensualdocente', '')
#                 try:
#                     os.remove(folder + generado + '.pdf')
#                 except Exception as ex:
#                     pass
#                 try:
#                     os.remove(folder + firmado + '.pdf')
#                 except Exception as ex:
#                     pass
#                 informe.delete()
#
#                 para = informe.distributivo.profesor.persona
#                 asunto = u"Firma invalida en informe mensual"
#                 observacion = f'Estimado/a Docente, hemos detectado una firma inválida en el proceso de entrega de su informe mensual, por ese motivo se ha eliminado el registro. Se solicita muy comedidamente que lo vuelva a generar y firmar desde el SGA, solo si su firma fue obtenida en una entidad diferente al SRI, REGISTRO CIVIL o BANCO CENTRAL. Si usted obtuvo su firma en una de las entidades antes mencionadas indicarlo a la Dirección de Innovación de Procesos Académicos mediante correo electrónico (direccion-innovacion@unemi.edu.ec) para que se socialice como puede entregar su informe mensual.'
#                 notificacion(asunto, observacion, para, None, 'pro_cronograma?action=vercumplimiento', informe.distributivo.pk, 1, 'sga', InformeMensualDocente, None)
#
#                 # print(f"-----{str(datos)}")
#                 informesinvalidos.append(informe.id)
#                 inicia += 1
#         print(len(informesinvalidos))
#         print('-----------------------------------')
#         print(informesinvalidos)
#     except Exception as ex:
#         print(ex)


# def aprobarrechazarpostulate():
#     cadena = ''
#     linea, excluidos, conexito = 0, 0, 0
#     try:
#         archivo_ = 'postulante_2_all'
#         # url_archivo = "{}/{}.xlsx".format(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), archivo_)
#         url_archivo = "{}/media/{}.xlsx".format(SITE_STORAGE, archivo_)
#         print('Leyendo....')
#         # wb = openpyxl.load_workbook(filename="{}/media/{}.xlsx".format(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), archivo_))
#         wb = openpyxl.load_workbook(filename=url_archivo)
#         ws = wb.get_sheet_by_name("faltan")
#         # worksheet = wb["Listado"]
#         worksheet = ws
#         lis_excluidos = []
#         print('Iniciando....')
#         linea_archivo = 1
#         ids_excluir_a_rechazar = []
#         for row in worksheet.iter_rows(min_row=0):
#             if linea >= 1:
#                 currentValues, cadena = [], ''
#                 for cell in row:
#                     cadena += str(cell.value) + ' '
#                     currentValues.append(str(cell.value))
#                 id_ = currentValues[0]
#                 cedula_ = currentValues[1]
#                 obstitulo_ = currentValues[4]
#                 aceptado_ = True if currentValues[3] == 'ACEPTADO' else False
#                 qsbasepersona = Persona.objects.filter(Q(cedula__icontains=cedula_)|Q(pasaporte__icontains=cedula_),status=True)
#                 if qsbasepersona.exists():
#                     persona_ = qsbasepersona.first()
#                     # qspostulacion = PersonaAplicarPartida.objects.filter(status=True, partida__convocatoria__vigente=True, partida__convocatoria__status=True, persona=persona_).order_by('partida')
#                     qspostulacion = PersonaAplicarPartida.objects.filter(status=True, id=id_, persona=persona_).order_by('partida')
#                     if qspostulacion.exists():
#                         postulacion_ = qspostulacion.first()
#
#                         print(f"{postulacion_} ({qspostulacion.count()})")
#                         posidiomas = PersonaIdiomaPartida.objects.filter(status=True, personapartida=postulacion_).order_by('id')
#                         postitulacion = PersonaFormacionAcademicoPartida.objects.filter(status=True, personapartida=postulacion_).order_by('id')
#                         posexperiencia = PersonaExperienciaPartida.objects.filter(status=True, personapartida=postulacion_).order_by('id')
#                         poscapacitacion = PersonaCapacitacionesPartida.objects.filter(status=True, personapartida=postulacion_, horas__gte=40).order_by('-horas')
#                         pospublicacion = PersonaPublicacionesPartida.objects.filter(status=True, personapartida=postulacion_).order_by('id')
#                         if not aceptado_:
#                             print(f"Rechazado")
#                             posidiomas.update(aceptado=False)
#                             postitulacion.update(aceptado=False)
#                             posexperiencia.update(aceptado=False)
#                             pospublicacion.update(aceptado=False)
#                             postulacion_.estado = 2
#                             postulacion_.obsgeneral = f"No cumple con la dedicación solicitada a tiempo completo, refleja que encuentra vinculado en otra institución."
#                             postulacion_.obsgeneral = obstitulo_ if obstitulo_ else ''
#                             postulacion_.obsgradoacademico = obstitulo_ if obstitulo_ else ''
#                         else:
#                             print(f"Aprobado")
#                             posidiomas.update(aceptado=False)
#                             postitulacion.update(aceptado=True)
#                             posexperiencia.update(aceptado=True)
#                             if poscapacitacion:
#                                 traerprimeracap_ = poscapacitacion.first()
#                                 traerprimeracap_.aceptado = True
#                                 traerprimeracap_.save()
#                             pospublicacion.update(aceptado=False)
#                             postulacion_.estado = 1
#                             postulacion_.obsgeneral = f"Cumple con lo solicitado."
#                         postulacion_.pgradoacademico = postulacion_.total_puntos_gradoacademico()
#                         postulacion_.pcapacitacion = postulacion_.total_puntos_capacitacion()
#                         postulacion_.pexpdocente = postulacion_.total_puntos_experiencia_docente()
#                         postulacion_.pexpadministrativa = postulacion_.total_puntos_experiencia_administrativo()
#                         postulacion_.nota_final_meritos = postulacion_.total_puntos()
#                         postulacion_.fecha_revision = datetime.now()
#                         postulacion_.calificada = True
#                         postulacion_.estado=4 if postulacion_.total_puntos() >=70 else 5
#                         postulacion_.save()
#                         calificacion = CalificacionPostulacion.objects.filter(status=True,postulacion=postulacion_).order_by('-id').first()
#                         if not calificacion:
#                             calificacion=CalificacionPostulacion(postulacion=postulacion_)
#                             calificacion.save()
#                         calificacion.pgradoacademico = postulacion_.total_puntos_gradoacademico()
#                         calificacion.obsgradoacademico = postulacion_.obsgradoacademico
#                         calificacion.pcapacitacion = postulacion_.total_puntos_capacitacion()
#                         calificacion.obscapacitacion = postulacion_.obscapacitacion
#                         calificacion.pexpdocente = postulacion_.total_puntos_experiencia_docente()
#                         calificacion.obsexperienciadoc = postulacion_.obsexperienciadoc
#                         calificacion.pexpadministrativa = postulacion_.total_puntos_experiencia_administrativo()
#                         calificacion.pexpadministrativa = postulacion_.pexpadministrativa
#                         calificacion.nota_final = postulacion_.total_puntos()
#                         calificacion.obsgeneral = postulacion_.obsgeneral
#                         calificacion.estado = postulacion_.estado
#                         calificacion.fecha_revision = datetime.now()
#                         calificacion.valida = True
#                         calificacion.save()
#
#                     else:
#                         lis_excluidos.append({'cedula': cedula_, 'obs': 'No existe postulación'})
#                 else:
#                     lis_excluidos.append({'cedula': cedula_, 'obs': 'No existe persona'})
#                     excluidos += 1
#                 print("Linea {}/? - Persona: {}".format(linea, cedula_))
#             linea += 1
#             linea_archivo += 1
#         print('Total Leidos con exito: {}'.format(conexito))
#         print('Total Excluidos: {}'.format(excluidos))
#         print(lis_excluidos)
#         # qsrechazados = PersonaAplicarPartida.objects.filter(status=True, partida__convocatoria__vigente=True, partida__convocatoria__status=True).exclude(id__in=ids_excluir_a_rechazar).order_by('partida')
#         # qsrechazados.update(fecha_revision=datetime.now(), estado=2, calificada=True, obsgeneral=f"No cumplen con los títulos requeridos en la convocatoria")
#         # wb.save(url_archivo)
#     except Exception as ex:
#         textoerror = '{} Linea:{} - Info: {} / Lectura: {}'.format(str(ex), sys.exc_info()[-1].tb_lineno, cadena, linea)
#         print(textoerror)
#
#
# aprobarrechazarpostulate()


#
#
# def aprobarrechazarpostulateapelacion():
#     cadena = ''
#     linea, excluidos, conexito = 0, 0, 0
#     try:
#         archivo_ = 'posapela'
#         # url_archivo = "{}/{}.xlsx".format(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), archivo_)
#         url_archivo = "{}/media/{}.xlsx".format(SITE_STORAGE, archivo_)
#         print('Leyendo....')
#         # wb = openpyxl.load_workbook(filename="{}/media/{}.xlsx".format(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), archivo_))
#         wb = openpyxl.load_workbook(filename=url_archivo)
#         ws = wb.get_sheet_by_name("postulantes")
#         # worksheet = wb["Listado"]
#         worksheet = ws
#         lis_excluidos = []
#         print('Iniciando....')
#         linea_archivo = 1
#         ids_excluir_a_rechazar = []
#         for row in worksheet.iter_rows(min_row=0):
#             if linea >= 1:
#                 currentValues, cadena = [], ''
#                 for cell in row:
#                     cadena += str(cell.value) + ' '
#                     currentValues.append(str(cell.value))
#                 id_ = currentValues[0]
#                 cedula_ = currentValues[3]
#                 rechazado = currentValues[8]
#                 obs = currentValues[9]
#                 qspostulacion = PersonaAplicarPartida.objects.filter(status=True, id=id_)
#                 if qspostulacion.exists():
#                     postulacion_ = qspostulacion.first()
#                     apelacion_ = postulacion_.traer_apelacion()
#                     print(f"{postulacion_} ({qspostulacion.count()})")
#                     if apelacion_:
#                         if rechazado == '0':
#                             print(f"Rechazado")
#                             apelacion_.estado = 2
#                             apelacion_.observacion = obs
#                             apelacion_.save()
#                         else:
#                             apelacion_.estado = 1
#                             apelacion_.observacion = obs
#                             apelacion_.save()
#                             calificacion = CalificacionPostulacion.objects.filter(status=True, postulacion=postulacion_).order_by('-id').first()
#                             if calificacion:
#                                 calificacion.valida = False
#                                 calificacion.save()
#                             posidiomas = PersonaIdiomaPartida.objects.filter(status=True, personapartida=postulacion_).order_by('id')
#                             postitulacion = PersonaFormacionAcademicoPartida.objects.filter(status=True, personapartida=postulacion_).order_by('id')
#                             posexperiencia = PersonaExperienciaPartida.objects.filter(status=True, personapartida=postulacion_).order_by('id')
#                             poscapacitacion = PersonaCapacitacionesPartida.objects.filter(status=True, personapartida=postulacion_).order_by('id')
#                             pospublicacion = PersonaPublicacionesPartida.objects.filter(status=True, personapartida=postulacion_).order_by('id')
#                             posidiomas.update(aceptado=True)
#                             postitulacion.update(aceptado=True)
#                             posexperiencia.update(aceptado=True)
#                             poscapacitacion.update(aceptado=True)
#                             pospublicacion.update(aceptado=True)
#                             postulacion_.pgradoacademico = postulacion_.total_puntos_gradoacademico()
#                             postulacion_.pcapacitacion = postulacion_.total_puntos_capacitacion()
#                             postulacion_.pexpdocente = postulacion_.total_puntos_experiencia_docente()
#                             postulacion_.pexpadministrativa = postulacion_.total_puntos_experiencia_administrativo()
#                             postulacion_.nota_final_meritos = postulacion_.total_puntos()
#                             postulacion_.fecha_revision = datetime.now()
#                             postulacion_.calificada = True
#                             postulacion_.estado = 1
#                             postulacion_.obsgeneral = obs
#                             postulacion_.save()
#                             calificacion = CalificacionPostulacion(postulacion=postulacion_)
#                             calificacion.pgradoacademico = postulacion_.total_puntos_gradoacademico()
#                             calificacion.obsgradoacademico = postulacion_.obsgradoacademico
#                             calificacion.pcapacitacion = postulacion_.total_puntos_capacitacion()
#                             calificacion.obscapacitacion = postulacion_.obscapacitacion
#                             calificacion.pexpdocente = postulacion_.total_puntos_experiencia_docente()
#                             calificacion.obsexperienciadoc = postulacion_.obsexperienciadoc
#                             calificacion.pexpadministrativa = postulacion_.total_puntos_experiencia_administrativo()
#                             calificacion.pexpadministrativa = postulacion_.pexpadministrativa
#                             calificacion.nota_final = postulacion_.total_puntos()
#                             calificacion.obsgeneral = postulacion_.obsgeneral
#                             calificacion.estado = postulacion_.estado
#                             calificacion.fecha_revision = datetime.now()
#                             calificacion.valida = True
#                             calificacion.save()
#                     else:
#                         lis_excluidos.append({'cedula': cedula_, 'obs': f'No existe apelación {id_}'})
#                         excluidos += 1
#                 else:
#                     lis_excluidos.append({'cedula': cedula_, 'obs': f'No existe postulación {id_}'})
#                     excluidos += 1
#                 print("Linea {}/? - Persona: {}".format(linea, cedula_))
#             linea += 1
#             linea_archivo += 1
#         print('Total Leidos con exito: {}'.format(conexito))
#         print('Total Excluidos: {}'.format(excluidos))
#         print(lis_excluidos)
#         # print(f"A Excluir: {len(ids_excluir_a_rechazar)}")
#         # qsrechazados = PersonaAplicarPartida.objects.filter(status=True, partida__convocatoria__vigente=True, partida__convocatoria__status=True).exclude(id__in=ids_excluir_a_rechazar).order_by('partida')
#         # qsrechazados.update(fecha_revision=datetime.now(), estado=2, calificada=True, obsgeneral=f"No cumplen con los títulos requeridos en la convocatoria")
#         # wb.save(url_archivo)
#     except Exception as ex:
#         textoerror = '{} Linea:{} - Info: {} / Lectura: {}'.format(str(ex), sys.exc_info()[-1].tb_lineno, cadena, linea)
#         print(textoerror)
#
#
# def modulossga():
#     # mis = ModuloCategorias.objects.create(nombre=f"General", prioridad=1)
#     # qsmodulos = Modulo.objects.all()
#     qsmodulos = Modulo.objects.filter(id__in=[28,
#                                               29,
#                                               30,
#                                               32,
#                                               33,
#                                               34,
#                                               48,
#                                               80,
#                                               98,
#                                               106,
#                                               132,
#                                               151,
#                                               197,
#                                               228,
#                                               270,
#                                               273,
#                                               275,
#                                               289,
#                                               299,
#                                               302,
#                                               359,
#                                               360,
#                                               371,
#                                               386,
#                                               390,
#                                               461,
#                                               369,
#                                               ])
#     for mod in qsmodulos:
#         print(mod)
#         mod.categorias.set([2])
#         mod.save()
#     qsmodulos = Modulo.objects.filter(id__in=[4,
#                                               44,
#                                               61,
#                                               119,
#                                               126,
#                                               179,
#                                               418,
#                                               430,
#                                               445,
#                                               459,
#                                               470,
#                                               479,
#                                               482,
#                                               452,
#                                               430, ])
#     for mod in qsmodulos:
#         print(mod)
#         mod.categorias.set([3])
#         mod.save()
#     qsmodulos = Modulo.objects.filter(id__in=[316,
#                                               367,
#                                               381,
#                                               423,
#                                               449,
#                                               477,
#                                               495,
#                                               510,
#                                               381, ])
#     for mod in qsmodulos:
#         print(mod)
#         mod.categorias.set([4])
#         mod.save()
#     qsmodulos = Modulo.objects.filter(id__in=[307,
#                                               334,
#                                               337,
#                                               346,
#                                               389,
#                                               472,
#                                               ])
#     for mod in qsmodulos:
#         print(mod)
#         mod.categorias.set([5])
#         mod.save()
#     qsmodulos = Modulo.objects.filter(id__in=[297,
#                                               403,
#                                               405,
#                                               406,
#                                               447,
#                                               493,
#                                               377, ])
#     for mod in qsmodulos:
#         print(mod)
#         mod.categorias.set([6])
#         mod.save()

# aprobarrechazarpostulateapelacion()

# def ADMMATRIZ():
#     cadena = ''
#     linea, excluidos, conexito = 0, 0, 0
#     try:
#         libre_origen = '/ADMMATRIZ.xls'
#         fuentecabecera = easyxf('font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
#         fuentenormal = easyxf('font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
#         output_folder = MEDIA_ROOT
#         output_folder = os.path.join(os.path.join(BASE_DIR))
#         # liborigen = xlrd.open_workbook(output_folder + libre_origen)
#         libdestino = xlwt.Workbook()
#         hojadestino = libdestino.add_sheet('HOJA1')
#         fil = 0
#         columnas = [
#             (u"PERIODO", 7000, 0),
#             (u"CODIGO_IES", 7000, 1),
#             (u"CODIGO_CARRERA", 7000, 2),
#             (u"CIUDAD_CARRERA", 7000, 3),
#             (u"TIPO_IDENTIFICACION", 7000, 4),
#             (u"IDENTIFICACION", 7000, 5),
#             (u"NOMBRES", 7000, 5),
#             (u"APELLIDOS", 7000, 5),
#             (u"SEXO", 7000, 6),
#             (u"PAIS_ORIGEN", 7000, 7),
#             (u"DISCAPACIDAD", 7000, 7),
#             (u"PORCENTAJE_DISCAPACIDAD", 7000, 7),
#             (u"NUMERO_CONADIS", 7000, 7),
#             (u"ETNIA", 7000, 7),
#             (u"NACIONALIDAD", 7000, 7),
#             (u"EMAIL_INSTITUCIONAL", 7000, 7),
#             (u"FECHA_INICIO_PRIMER_NIVEL", 7000, 7),
#             (u"FECHA_INGRESO_CONVALIDACION", 7000, 7),
#             (u"PAIS_RESIDENCIA", 7000, 7),
#             (u"PROVINCIA_RESIDENCIA", 7000, 7),
#             (u"CANTON_RESIDENCIA", 7000, 7),
#             (u"TIPO_COLEGIO", 7000, 7),
#             (u"POLITICA_CUOTA", 7000, 7),
#             (u"CARRERA", 7000, 7),
#         ]
#         for col_num in range(len(columnas)):
#             hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
#             hojadestino.col(col_num).width = columnas[col_num][1]
#         qsmatriculas = Matricula.objects.filter(status=True, nivel__periodo__in=(126)).order_by('nivel__periodo')
#         fila = 1
#         totmaterias = len(qsmatriculas)
#         for sm in qsmatriculas:
#             print('{}/{} - {}'.format(totmaterias, fila, sm))
#             persona_ = sm.inscripcion.persona
#             inscripcion_ = sm.inscripcion
#             ficha_ = persona_.mi_ficha()
#             matricula = sm
#             hojadestino.write(fila, 0, matricula.nivel.periodo.__str__(), fuentenormal)
#             hojadestino.write(fila, 1, '59', fuentenormal)
#             hojadestino.write(fila, 2, '00098', fuentenormal)
#             hojadestino.write(fila, 3, 'MILAGRO', fuentenormal)
#             tp_identificacion = ''
#             if persona_.tipo_identificacion() == 'C':
#                 tp_identificacion = 'CEDULA'
#             elif persona_.tipo_identificacion() == 'P':
#                 tp_identificacion = 'PASAPORTE'
#             elif persona_.tipo_identificacion() == 'R':
#                 tp_identificacion = 'RUC'
#             hojadestino.write(fila, 4, tp_identificacion, fuentenormal)
#             hojadestino.write(fila, 5, persona_.identificacion(), fuentenormal)
#             hojadestino.write(fila, 6, persona_.nombres, fuentenormal)
#             hojadestino.write(fila, 7, f"{persona_.apellido1} {persona_.apellido2}", fuentenormal)
#             hojadestino.write(fila, 8, persona_.sexo.__str__() if persona_.sexo else '', fuentenormal)
#             hojadestino.write(fila, 9, persona_.paisnacimiento.__str__() if persona_.paisnacimiento else '', fuentenormal)
#             perfil_ = persona_.mi_perfil()
#             if perfil_.tienediscapacidad:
#                 hojadestino.write(fila, 10, perfil_.tipodiscapacidad.nombre if perfil_.tipodiscapacidad else '', fuentenormal)
#                 hojadestino.write(fila, 11, perfil_.porcientodiscapacidad, fuentenormal)
#                 hojadestino.write(fila, 12, perfil_.carnetdiscapacidad, fuentenormal)
#                 hojadestino.write(fila, 13, perfil_.raza.__str__() if perfil_.raza else '', fuentenormal)
#             else:
#                 hojadestino.write(fila, 10, '', fuentenormal)
#                 hojadestino.write(fila, 11, '', fuentenormal)
#                 hojadestino.write(fila, 12, '', fuentenormal)
#                 hojadestino.write(fila, 13, '', fuentenormal)
#             hojadestino.write(fila, 14, persona_.nacionalidad, fuentenormal)
#             hojadestino.write(fila, 15, persona_.emailinst, fuentenormal)
#             # mprimernivel = Matricula.objects.filter(status=True, inscripcion__persona=persona_, inscripcion__coordinacion__in=[9,]).order_by('id')
#             # if mprimernivel:
#             #     fprimera = str(mprimernivel.first().fecha)
#             #     hojadestino.write(fila, 16, fprimera, fuentenormal)
#             # else:
#             #     hojadestino.write(fila, 16, '', fuentenormal)
#             hojadestino.write(fila, 16, str(sm.fecha) if sm.fecha else str(sm.nivel.periodo.inicio), fuentenormal)
#             hojadestino.write(fila, 17, '', fuentenormal)
#             hojadestino.write(fila, 18, persona_.pais.__str__(), fuentenormal)
#             hojadestino.write(fila, 19, persona_.provincia.__str__(), fuentenormal)
#             hojadestino.write(fila, 20, persona_.canton.__str__(), fuentenormal)
#             qstitulo = persona_.titulacion_set.filter(status=True).order_by('-id').first()
#             if qstitulo:
#                 hojadestino.write(fila, 21, qstitulo.colegio.get_tipo_display() if qstitulo.colegio else '', fuentenormal)
#             else:
#                 hojadestino.write(fila, 21, '', fuentenormal)
#             hojadestino.write(fila, 22, '', fuentenormal)
#             hojadestino.write(fila, 23, inscripcion_.carrera.__str__(), fuentenormal)
#             fila += 1
#
#         libdestino.save(output_folder + libre_origen)
#         print(output_folder + libre_origen)
#         print("Proceso finalizado. . .")
#     except Exception as ex:
#         msg = ex.__str__()
#
#         textoerror = '{} Linea:{}'.format(str(ex), sys.exc_info()[-1].tb_lineno)
#         print(textoerror)
#         print(msg)


# def PREGRADO_MATRIZ_1S_2022():
#     try:
#         fuentecabecera = easyxf('font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
#         fuentenormal = easyxf('font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
#         output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media', 'reportes', '2023', '05', '23', ''))
#         os.makedirs(output_folder, exist_ok=True)
#         nombre = "matriz_pregrado_1s_2022_" + datetime.now().strftime('%Y%m%d_%H%M%S') + ".xls"
#         filename = os.path.join(output_folder, nombre)
#         libdestino = xlwt.Workbook()
#         hojadestino = libdestino.add_sheet('HOJA1')
#         fil = 0
#         columnas = [
#             (u"PERIODO", 7000),
#             (u"CODIGO_IES", 7000),
#             (u"ID_CARRERA", 7000),
#             (u"CODIGO_CARRERA", 7000),
#             (u"NOMBRE_CARRERA", 7000),
#             (u"CIUDAD_CARRERA", 7000),
#             (u"TIPO_IDENTIFICACION", 7000),
#             (u"IDENTIFICACION", 7000),
#             (u"NOMBRES", 7000),
#             (u"APELLIDOS", 7000),
#             (u"SEXO", 7000),
#             (u"FECHA_MATRICULA", 7000),
#             (u"FECHA_CONVALIDACION", 7000),
#             (u"TOTAL_CREDITOS_APROBADOS", 7000),
#             (u"CREDITOS_APROBADOS", 7000),
#             (u"NIVEL_ACADEMICO", 7000),
#             (u"NUM_MATERIAS_SEGUNDA_MATRICULA", 7000),
#             (u"NUM_MATERIAS_TERCERA_MATRICULA", 7000),
#             (u"PERDIDA_GRATUIDAD", 7000),
#             (u"INGRESO_TOTAL_HOGAR", 7000),
#             (u"TOTAL_HORAS_APROBADAS", 7000),
#             (u"HORAS_APROBADAS_PERIODO", 7000),
#             (u"MONTO_AYUDA_ECONOMICA", 7000),
#             (u"MONTO_CREDITO_EDUCATIVO", 7000),
#             (u"ESTADO", 7000),
#         ]
#         for col_num in range(len(columnas)):
#             hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
#             hojadestino.col(col_num).width = columnas[col_num][1]
#         eMatriculas = Matricula.objects.filter(status=True, retiradomatricula=False,  nivel__periodo_id=126, inscripcion__coordinacion__id__in=[2, 3]).order_by('nivel__periodo')
#         # eMatriculas = eMatriculas[:10]
#         fila = 1
#         total = len(eMatriculas.values("id"))
#         for eMatricula in eMatriculas:
#             print('{}/{} - {}'.format(total, fila, eMatricula))
#             ePersona = eMatricula.inscripcion.persona
#             ePeriodo = eMatricula.nivel.periodo
#             eInscripcion = eMatricula.inscripcion
#             eCarrera = eInscripcion.carrera
#             # eFicha = ePersona.mi_ficha()
#             eMalla = eInscripcion.mi_malla()
#             hojadestino.write(fila, 0, str(ePeriodo.__str__()).strip(), fuentenormal) #PERIODO
#             hojadestino.write(fila, 1, '1024', fuentenormal) #CODIGO_IES
#             hojadestino.write(fila, 2, str(eCarrera.id), fuentenormal) #ID_CARRERA
#             codigo_carrera = ''
#             if eMalla.codigo:
#                 codigo_carrera = eMalla.codigo
#             elif eCarrera.codigo:
#                 codigo_carrera = eCarrera.codigo
#             hojadestino.write(fila, 3, f"{codigo_carrera}".strip(), fuentenormal) #CODIGO_CARRERA
#             hojadestino.write(fila, 4, str(eCarrera.nombre_completo()).strip(), fuentenormal) #NOMBRE_CARRERA
#             hojadestino.write(fila, 5, "MILAGRO", fuentenormal) #CIUDAD_CARRERA
#             hojadestino.write(fila, 6, ePersona.tipo_identificacion_completo(), fuentenormal) #TIPO_IDENTIFICACION
#             hojadestino.write(fila, 7, ePersona.identificacion(), fuentenormal) #IDENTIFICACION
#             hojadestino.write(fila, 8, ePersona.nombres, fuentenormal) #NOMBRES
#             hojadestino.write(fila, 9,  f"{ePersona.apellido1} {ePersona.apellido2}".strip(), fuentenormal) #APELLIDOS
#             hojadestino.write(fila, 10,  f"{ePersona.sexo.__str__() if ePersona.sexo else ''}".strip(), fuentenormal) #SEXO
#             hojadestino.write(fila, 11,  str(eMatricula.fecha), fuentenormal) #FECHA_MATRICULA
#             hojadestino.write(fila, 12,  str(eInscripcion.fechainicioconvalidacion) if eInscripcion.fechainicioconvalidacion else '', fuentenormal) #FECHA_CONVALIDACION
#             total_creditos_aprobados = null_to_numeric(eInscripcion.recordacademico_set.filter(valida=True, status=True, aprobada=True).aggregate(creditos=Sum('creditos'))['creditos'], 2)
#             hojadestino.write(fila, 13,  str(total_creditos_aprobados), fuentenormal) #TOTAL_CREDITOS_APROBADOS
#             creditos_aprobados = null_to_numeric(MateriaAsignada.objects.filter(matricula=eMatricula, retiramateria=False, status=True, estado_id=NOTA_ESTADO_APROBADO, materia__asignaturamalla__isnull=False).aggregate(creditos=Sum('materia__asignaturamalla__creditos'))['creditos'], 2)
#             hojadestino.write(fila, 14,  str(creditos_aprobados), fuentenormal) #CREDITOS_APROBADOS
#             hojadestino.write(fila, 15,  eMatricula.nivelmalla.nombre, fuentenormal) #NIVEL_ACADEMICO
#             num_materias_segunda_matricula = MateriaAsignada.objects.values("id").filter(matricula=eMatricula, retiramateria=False, status=True, matriculas=2, materia__asignaturamalla__isnull=False).count()
#             hojadestino.write(fila, 16,  str(num_materias_segunda_matricula), fuentenormal) #NUM_MATERIAS_SEGUNDA_MATRICULA
#             num_materias_tercera_matricula = MateriaAsignada.objects.values("id").filter(matricula=eMatricula, retiramateria=False, status=True, matriculas__gte=3, materia__asignaturamalla__isnull=False).count()
#             hojadestino.write(fila, 17,  str(num_materias_tercera_matricula), fuentenormal) #NUM_MATERIAS_TERCERA_MATRICULA
#             hojadestino.write(fila, 18,  'SI' if eInscripcion.estado_gratuidad == 3 else 'NO', fuentenormal) #PERDIDA_GRATUIDAD
#             hojadestino.write(fila, 19,  'NO REGISTRA', fuentenormal) #INGRESO_TOTAL_HOGAR
#             total_horas_aprobados = null_to_numeric(eInscripcion.recordacademico_set.filter(valida=True, status=True, aprobada=True).aggregate(horas=Sum('horas'))['horas'], 0)
#             hojadestino.write(fila, 20,  str(total_horas_aprobados), fuentenormal) #TOTAL_HORAS_APROBADAS
#             horas_aprobados = null_to_numeric(MateriaAsignada.objects.filter(matricula=eMatricula, retiramateria=False, status=True, estado_id=NOTA_ESTADO_APROBADO, materia__asignaturamalla__isnull=False).aggregate(horas=Sum('materia__asignaturamalla__horas'))['horas'], 0)
#             hojadestino.write(fila, 21, str(horas_aprobados), fuentenormal) #HORAS_APROBADAS_PERIODO
#             hojadestino.write(fila, 22, '0', fuentenormal) #MONTO_AYUDA_ECONOMICA
#             hojadestino.write(fila, 23, '0', fuentenormal) #MONTO_CREDITO_EDUCATIVO
#             hojadestino.write(fila, 24, 'NO APLICA', fuentenormal) #ESTADO
#             fila += 1
#
#         libdestino.save(filename)
#         print(filename)
#         print("Proceso finalizado. . .")
#     except Exception as ex:
#         msg = ex.__str__()
#
#         textoerror = '{} Linea:{}'.format(str(ex), sys.exc_info()[-1].tb_lineno)
#         print(textoerror)
#         print(msg)


# def PREGRADO_MATRIZ_2S_2022():
#     try:
#         fuentecabecera = easyxf('font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
#         fuentenormal = easyxf('font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
#         output_folder = os.path.join(os.path.join(SITE_STORAGE, 'media', 'reportes', '2023', '05', '23', ''))
#         os.makedirs(output_folder, exist_ok=True)
#         nombre = "matriz_pregrado_2s_2022_" + datetime.now().strftime('%Y%m%d_%H%M%S') + ".xls"
#         filename = os.path.join(output_folder, nombre)
#         libdestino = xlwt.Workbook()
#         hojadestino = libdestino.add_sheet('HOJA1')
#         fil = 0
#         columnas = [
#             (u"PERIODO", 7000),
#             (u"CODIGO_IES", 7000),
#             (u"ID_CARRERA", 7000),
#             (u"CODIGO_CARRERA", 7000),
#             (u"NOMBRE_CARRERA", 7000),
#             (u"CIUDAD_CARRERA", 7000),
#             (u"TIPO_IDENTIFICACION", 7000),
#             (u"IDENTIFICACION", 7000),
#             (u"NOMBRES", 7000),
#             (u"APELLIDOS", 7000),
#             (u"SEXO", 7000),
#             (u"FECHA_MATRICULA", 7000),
#             (u"FECHA_CONVALIDACION", 7000),
#             (u"TOTAL_CREDITOS_APROBADOS", 7000),
#             (u"CREDITOS_APROBADOS", 7000),
#             (u"NIVEL_ACADEMICO", 7000),
#             (u"NUM_MATERIAS_SEGUNDA_MATRICULA", 7000),
#             (u"NUM_MATERIAS_TERCERA_MATRICULA", 7000),
#             (u"PERDIDA_GRATUIDAD", 7000),
#             (u"INGRESO_TOTAL_HOGAR", 7000),
#             (u"TOTAL_HORAS_APROBADAS", 7000),
#             (u"HORAS_APROBADAS_PERIODO", 7000),
#             (u"MONTO_AYUDA_ECONOMICA", 7000),
#             (u"MONTO_CREDITO_EDUCATIVO", 7000),
#             (u"ESTADO", 7000),
#         ]
#         for col_num in range(len(columnas)):
#             hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
#             hojadestino.col(col_num).width = columnas[col_num][1]
#         eMatriculas = Matricula.objects.filter(status=True, retiradomatricula=False,  nivel__periodo_id=153, inscripcion__coordinacion__id__in=[2, 3]).order_by('nivel__periodo')
#         # eMatriculas = eMatriculas[:10]
#         fila = 1
#         total = len(eMatriculas.values("id"))
#         for eMatricula in eMatriculas:
#             print('{}/{} - {}'.format(total, fila, eMatricula))
#             ePersona = eMatricula.inscripcion.persona
#             ePeriodo = eMatricula.nivel.periodo
#             eInscripcion = eMatricula.inscripcion
#             eCarrera = eInscripcion.carrera
#             # eFicha = ePersona.mi_ficha()
#             eMalla = eInscripcion.mi_malla()
#             hojadestino.write(fila, 0, str(ePeriodo.__str__()).strip(), fuentenormal) #PERIODO
#             hojadestino.write(fila, 1, '1024', fuentenormal) #CODIGO_IES
#             hojadestino.write(fila, 2, str(eCarrera.id), fuentenormal) #ID_CARRERA
#             codigo_carrera = ''
#             if eMalla.codigo:
#                 codigo_carrera = eMalla.codigo
#             elif eCarrera.codigo:
#                 codigo_carrera = eCarrera.codigo
#             hojadestino.write(fila, 3, f"{codigo_carrera}".strip(), fuentenormal) #CODIGO_CARRERA
#             hojadestino.write(fila, 4, str(eCarrera.nombre_completo()).strip(), fuentenormal) #NOMBRE_CARRERA
#             hojadestino.write(fila, 5, "MILAGRO", fuentenormal) #CIUDAD_CARRERA
#             hojadestino.write(fila, 6, ePersona.tipo_identificacion_completo(), fuentenormal) #TIPO_IDENTIFICACION
#             hojadestino.write(fila, 7, ePersona.identificacion(), fuentenormal) #IDENTIFICACION
#             hojadestino.write(fila, 8, ePersona.nombres, fuentenormal) #NOMBRES
#             hojadestino.write(fila, 9,  f"{ePersona.apellido1} {ePersona.apellido2}".strip(), fuentenormal) #APELLIDOS
#             hojadestino.write(fila, 10,  f"{ePersona.sexo.__str__() if ePersona.sexo else ''}".strip(), fuentenormal) #SEXO
#             hojadestino.write(fila, 11,  str(eMatricula.fecha), fuentenormal) #FECHA_MATRICULA
#             hojadestino.write(fila, 12,  str(eInscripcion.fechainicioconvalidacion) if eInscripcion.fechainicioconvalidacion else '', fuentenormal) #FECHA_CONVALIDACION
#             total_creditos_aprobados = null_to_numeric(eInscripcion.recordacademico_set.filter(valida=True, status=True, aprobada=True).aggregate(creditos=Sum('creditos'))['creditos'], 2)
#             hojadestino.write(fila, 13,  str(total_creditos_aprobados), fuentenormal) #TOTAL_CREDITOS_APROBADOS
#             creditos_aprobados = null_to_numeric(MateriaAsignada.objects.filter(matricula=eMatricula, retiramateria=False, status=True, estado_id=NOTA_ESTADO_APROBADO, materia__asignaturamalla__isnull=False).aggregate(creditos=Sum('materia__asignaturamalla__creditos'))['creditos'], 2)
#             hojadestino.write(fila, 14,  str(creditos_aprobados), fuentenormal) #CREDITOS_APROBADOS
#             hojadestino.write(fila, 15,  eMatricula.nivelmalla.nombre, fuentenormal) #NIVEL_ACADEMICO
#             num_materias_segunda_matricula = MateriaAsignada.objects.values("id").filter(matricula=eMatricula, retiramateria=False, status=True, matriculas=2, materia__asignaturamalla__isnull=False).count()
#             hojadestino.write(fila, 16,  str(num_materias_segunda_matricula), fuentenormal) #NUM_MATERIAS_SEGUNDA_MATRICULA
#             num_materias_tercera_matricula = MateriaAsignada.objects.values("id").filter(matricula=eMatricula, retiramateria=False, status=True, matriculas__gte=3, materia__asignaturamalla__isnull=False).count()
#             hojadestino.write(fila, 17,  str(num_materias_tercera_matricula), fuentenormal) #NUM_MATERIAS_TERCERA_MATRICULA
#             hojadestino.write(fila, 18,  'SI' if eInscripcion.estado_gratuidad == 3 else 'NO', fuentenormal) #PERDIDA_GRATUIDAD
#             hojadestino.write(fila, 19,  'NO REGISTRA', fuentenormal) #INGRESO_TOTAL_HOGAR
#             total_horas_aprobados = null_to_numeric(eInscripcion.recordacademico_set.filter(valida=True, status=True, aprobada=True).aggregate(horas=Sum('horas'))['horas'], 0)
#             hojadestino.write(fila, 20,  str(total_horas_aprobados), fuentenormal) #TOTAL_HORAS_APROBADAS
#             horas_aprobados = null_to_numeric(MateriaAsignada.objects.filter(matricula=eMatricula, retiramateria=False, status=True, estado_id=NOTA_ESTADO_APROBADO, materia__asignaturamalla__isnull=False).aggregate(horas=Sum('materia__asignaturamalla__horas'))['horas'], 0)
#             hojadestino.write(fila, 21, str(horas_aprobados), fuentenormal) #HORAS_APROBADAS_PERIODO
#             hojadestino.write(fila, 22, '0', fuentenormal) #MONTO_AYUDA_ECONOMICA
#             hojadestino.write(fila, 23, '0', fuentenormal) #MONTO_CREDITO_EDUCATIVO
#             hojadestino.write(fila, 24, 'NO APLICA', fuentenormal) #ESTADO
#             fila += 1
#
#         libdestino.save(filename)
#         print(filename)
#         print("Proceso finalizado. . .")
#     except Exception as ex:
#         msg = ex.__str__()
#
#         textoerror = '{} Linea:{}'.format(str(ex), sys.exc_info()[-1].tb_lineno)
#         print(textoerror)
#         print(msg)


#
#
# def asignargrupo():
#     try:
#         qslist = PartidaTribunal.objects.filter(status=True, tipo=1, partida__convocatoria__vigente=True)
#         for i in qslist:
#             print(f"{i}")
#             persona_ = i.persona
#             user_ = persona_.usuario
#             if user_:
#                 if i.tipo == 1:
#                     if not user_.groups.filter(id=360).exists():
#                         user_.groups.add(360)
#                         print(f"-------Asignado")
#                 else:
#                     if not user_.groups.filter(id=362).exists():
#                         user_.groups.add(362)
#                         print(f"-------Asignado")
#     except Exception as ex:
#         print(ex)
#
#
# def reportedeudas31enero():
#     cadena = ''
#     linea, excluidos, conexito = 0, 0, 0
#     try:
#         libre_origen = '/31PAGOS.xls'
#         fuentecabecera = easyxf('font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
#         fuentenormal = easyxf('font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
#         output_folder = MEDIA_ROOT
#         output_folder = os.path.join(os.path.join(BASE_DIR))
#         # liborigen = xlrd.open_workbook(output_folder + libre_origen)
#         libdestino = xlwt.Workbook()
#         hojadestino = libdestino.add_sheet('HOJA1')
#         fil = 0
#         columnas = [
#             (u"PERIODO", 7000, 0),
#             (u"COORDINACION", 7000, 1),
#             (u"CARRERA", 7000, 1),
#             (u"NIVEL", 7000, 2),
#             (u"CEDULA", 7000, 3),
#             (u"PERSONA", 7000, 4),
#             (u"RUBRO", 7000, 5),
#             (u"TOTAL", 7000, 6),
#             (u"SALDO", 7000, 7),
#         ]
#         for col_num in range(len(columnas)):
#             hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
#             hojadestino.col(col_num).width = columnas[col_num][1]
#
#         qspagos = Pago.objects.filter(status=True, rubro__matricula__nivel__periodo__id=153, fecha_creacion__gte='2023-01-31').values_list('rubro__id', flat=True)
#         qsrubros = Rubro.objects.filter(status=True, id__in=list(qspagos))
#         fila = 1
#         totmaterias = len(qsrubros)
#         for sm in qsrubros:
#             print('{}/{} - {}'.format(totmaterias, fila, sm))
#             persona_ = sm.matricula.inscripcion.persona
#             inscripcion_ = sm.matricula.inscripcion
#             matricula_ = sm.matricula
#             hojadestino.write(fila, 0, matricula_.nivel.periodo.__str__(), fuentenormal)
#             hojadestino.write(fila, 1, inscripcion_.coordinacion.nombre, fuentenormal)
#             hojadestino.write(fila, 2, inscripcion_.carrera.nombre, fuentenormal)
#             hojadestino.write(fila, 3, matricula_.nivelmalla.nombre, fuentenormal)
#             hojadestino.write(fila, 4, persona_.cedula, fuentenormal)
#             hojadestino.write(fila, 5, persona_.nombre_completo(), fuentenormal)
#             hojadestino.write(fila, 6, sm.nombre, fuentenormal)
#             hojadestino.write(fila, 7, sm.valortotal, fuentenormal)
#             hojadestino.write(fila, 8, sm.saldo, fuentenormal)
#             qsmodelo = sm.pago_set.filter(status=True).order_by('fecha_creacion')
#             col_nums = 9
#             for modelo in qsmodelo:
#                 hojadestino.write(fila, col_nums, str(modelo.fecha_creacion.date()) if modelo else '', fuentenormal)
#                 hojadestino.write(fila, col_nums+1, modelo.valortotal if modelo else '', fuentenormal)
#                 col_nums += 2
#             fila += 1
#
#         libdestino.save(output_folder + libre_origen)
#         print(output_folder + libre_origen)
#         print("Proceso finalizado. . .")
#     except Exception as ex:
#         msg = ex.__str__()
#
#         textoerror = '{} Linea:{}'.format(str(ex), sys.exc_info()[-1].tb_lineno)
#         print(textoerror)
#         print(msg)
#
#
# def reportetic():
#     cadena = ''
#     linea, excluidos, conexito = 0, 0, 0
#     try:
#         libre_origen = '/TICMATERIAS.xls'
#         fuentecabecera = easyxf('font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
#         fuentenormal = easyxf('font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
#         output_folder = MEDIA_ROOT
#         output_folder = os.path.join(os.path.join(BASE_DIR))
#         # liborigen = xlrd.open_workbook(output_folder + libre_origen)
#         libdestino = xlwt.Workbook()
#         hojadestino = libdestino.add_sheet('HOJA1')
#         fil = 0
#         columnas = [
#             (u"PERIODO", 7000, 0),
#             (u"ASIGNATURA", 7000, 1),
#             (u"CARRERA", 7000, 2),
#             (u"NIVEL", 7000, 3),
#             (u"MODALIDAD", 7000, 4),
#             (u"CEDULA", 7000, 5),
#             (u"PERSONA", 7000, 6),
#             (u"EDAD", 7000, 7),
#             (u"ESTADO CIVIL", 7000, 8),
#             (u"SEXO", 7000, 9),
#             (u"COLEGIO", 7000, 10),
#             (u"TIPO COLEGIO", 7000, 11),
#             (u"F. BACHILLER", 7000, 12),
#             (u"ACCESOS IES PUNTAJE", 7000, 13),
#             (u"F. INGRESO PRIMER NIVEL", 7000, 14),
#             (u"PROBLEMA DE APRENDIZAJE", 7000, 15),
#             (u"TIENE DISCAPACIDAD", 7000, 16),
#             (u"DISCAPACIDAD", 7000, 17),
#             (u"TIENE TRABAJO", 7000, 18),
#             (u"LUGAR TRABAJO", 7000, 19),
#             (u"ES CABEZA DE FAMILIA", 7000, 20),
#             (u"OCUPACION CABEZA DE FAMILIA", 7000, 21),
#             (u"TOTAL INGRESOS FAMILIARES", 7000, 22),
#             (u"ESTADO DE SALUD", 7000, 23),
#             (u"ESTUDIA OTRA CARRERA", 7000, 24),
#             (u"TIPO VIVIENDA", 7000, 25),
#             (u"VIVIENDA", 7000, 26),
#             (u"HORAS USO INTERNET", 7000, 27),
#             (u"RECIBE AYUDA FAMILIAR", 7000, 28),
#             (u"RECIBE BECA", 7000, 29),
#             (u"HIJOS NUMERO", 7000, 30),
#             (u"TIENE FAMILIAR ENFERMEDAD", 7000, 31),
#             (u"TIENE FAMILIAR DISCAPACIDAD", 7000, 32),
#             (u"ESTADO MATERIA", 7000, 33),
#             (u"NOTA ADMISION FINAL", 7000, 34),
#             (u"ASISTENCIA FINAL", 7000, 35),
#             (u"NOTA FINAL", 7000, 36),
#             (u"TOTAL CREDITOS DE MATRICULA", 7000, 37),
#             (u"TOTAL CREDITOS MATERIA", 7000, 38),
#             (u"NUM. VECES CURSANDO", 7000, 39),
#         ]
#         for col_num in range(len(columnas)):
#             hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
#             hojadestino.col(col_num).width = columnas[col_num][1]
#
#         qsmateria = MateriaAsignada.objects.filter(status=True, materia__asignatura__id=3309)
#         qsmodelo = qsmateria[0].evaluaciongenerica_set.filter(status=True).order_by('detallemodeloevaluativo__orden')
#         col_num = 40
#         for modelo in qsmodelo:
#             hojadestino.write(fil, col_num, modelo.detallemodeloevaluativo.nombre, fuentecabecera)
#             hojadestino.col(col_num).width = 7000
#             col_num += 1
#
#         fila = 1
#         totmaterias = len(qsmateria)
#         for sm in qsmateria:
#             print('{}/{} - {}'.format(totmaterias, fila, sm))
#             persona_ = sm.matricula.inscripcion.persona
#             inscripcion_ = sm.matricula.inscripcion
#             ficha_ = persona_.mi_ficha()
#             matricula_ = sm.matricula
#             qstitulo = persona_.titulacion_set.filter(status=True).order_by('-id').first()
#             personainscripciones_ = Inscripcion.objects.values('id').filter(status=True, persona=persona_, activo=True).exclude(coordinacion__in=[9, 7])
#             becaasignacion_ = inscripcion_.becasolicitud_set.values('id').filter(status=True, estado__in=[29, 30]).exists()
#             familiares_ = persona_.personadatosfamiliares_set.filter(status=True)
#             numhijos = familiares_.filter(parentesco__in=[14,11]).count()
#             numdiscapacidad = familiares_.filter(tienediscapacidad=True).count()
#             hojadestino.write(fila, 0, sm.matricula.nivel.periodo.__str__(), fuentenormal)
#             hojadestino.write(fila, 1, sm.materia.asignatura.__str__(), fuentenormal)
#             hojadestino.write(fila, 2, sm.materia.asignaturamalla.malla.carrera.nombre, fuentenormal)
#             hojadestino.write(fila, 3, sm.materia.asignaturamalla.nivelmalla.nombre, fuentenormal)
#             hojadestino.write(fila, 4, inscripcion_.modalidad.__str__(), fuentenormal)
#             hojadestino.write(fila, 5, persona_.cedula, fuentenormal)
#             hojadestino.write(fila, 6, persona_.nombre_completo(), fuentenormal)
#             hojadestino.write(fila, 7, persona_.edad(), fuentenormal)
#             if persona_.datos_extension():
#                 hojadestino.write(fila, 8, persona_.estado_civil_descripcion() if persona_.datos_extension().estadocivil else '', fuentenormal)
#             else:
#                 hojadestino.write(fila, 8, '', fuentenormal)
#             hojadestino.write(fila, 9, persona_.sexo.__str__() if persona_.sexo else '', fuentenormal)
#             if qstitulo:
#                 hojadestino.write(fila, 10, qstitulo.colegio.nombre if qstitulo.colegio else '', fuentenormal)
#                 hojadestino.write(fila, 11, qstitulo.colegio.get_tipo_display() if qstitulo.colegio else '', fuentenormal)
#                 hojadestino.write(fila, 12, str(qstitulo.fechaobtencion) if qstitulo.colegio else '', fuentenormal)
#             else:
#                 hojadestino.write(fila, 10, '', fuentenormal)
#                 hojadestino.write(fila, 11, '', fuentenormal)
#                 hojadestino.write(fila, 12, '', fuentenormal)
#             hojadestino.write(fila, 13, inscripcion_.puntajesenescyt, fuentenormal) #iespuntaje
#             hojadestino.write(fila, 14, str(inscripcion_.fechainicioprimernivel) if qstitulo else '', fuentenormal)
#             hojadestino.write(fila, 15, '', fuentenormal) #problemaaprendizaje
#             discapacidad = persona_.mi_perfil()
#             hojadestino.write(fila, 16, discapacidad.tienediscapacidad, fuentenormal)
#             hojadestino.write(fila, 17, discapacidad.tipodiscapacidad.nombre if discapacidad.tipodiscapacidad else '', fuentenormal)
#             hojadestino.write(fila, 18, persona_.situacion_laboral().disponetrabajo if persona_.situacion_laboral() else '', fuentenormal)
#             hojadestino.write(fila, 19, persona_.situacion_laboral().lugartrabajo if persona_.situacion_laboral() else '', fuentenormal)
#             hojadestino.write(fila, 20, ficha_.escabezafamilia, fuentenormal)
#             hojadestino.write(fila, 21, ficha_.ocupacionjefehogar.nombre if ficha_.ocupacionjefehogar else '', fuentenormal)
#             hojadestino.write(fila, 22, persona_.total_ingresos_familiares())
#             hojadestino.write(fila, 23, ficha_.get_estadogeneral_display(), fuentenormal) #salud
#             hojadestino.write(fila, 24, personainscripciones_.count(), fuentenormal) #segundacarrera
#             hojadestino.write(fila, 25, ficha_.tipovivienda.__str__() if ficha_ else '', fuentenormal)
#             hojadestino.write(fila, 26, '', fuentenormal) #vivienda
#             hojadestino.write(fila, 27, ficha_.val_usainternetseism, fuentenormal) #horausointernet
#             hojadestino.write(fila, 28, ficha_.esdependiente, fuentenormal) #recibeayuda
#             hojadestino.write(fila, 29, becaasignacion_, fuentenormal) #recibebeca
#             hojadestino.write(fila, 30, numhijos, fuentenormal) #hijosnumero
#             hojadestino.write(fila, 31, '', fuentenormal) #familiarenfermedad
#             hojadestino.write(fila, 32, numdiscapacidad, fuentenormal) #familiardiscapacidad
#             hojadestino.write(fila, 33, sm.estado.nombre if sm.estado else '', fuentenormal)
#             notaadmisionfinal = 0
#             qsmatriculaadm = Matricula.objects.filter(status=True, inscripcion__persona=persona_, inscripcion__coordinacion__id=9).order_by('-id')
#             if qsmatriculaadm.exists():
#                 notaadmisionfinal = qsmatriculaadm.first().promedio_nota()
#             hojadestino.write(fila, 34, notaadmisionfinal, fuentenormal)
#             hojadestino.write(fila, 35, sm.asistenciafinal, fuentenormal)
#             hojadestino.write(fila, 36, sm.notafinal, fuentenormal)
#             totcreditos = MateriaAsignada.objects.filter(status=True, matricula=matricula_, retiramateria=False).aggregate(numcreditos=Coalesce(Sum(F('materia__creditos'), output_field=IntegerField()), 0)).get('numcreditos')
#             hojadestino.write(fila, 37, totcreditos, fuentenormal)
#             hojadestino.write(fila, 38, sm.materia.creditos, fuentenormal)
#             numcursandomateria = MateriaAsignada.objects.values('id').filter(status=True, matricula__inscripcion=inscripcion_, materia__asignatura__id=3309).count()
#             hojadestino.write(fila, 39, numcursandomateria, fuentenormal)
#             qsmodelo = sm.evaluaciongenerica_set.filter(status=True).order_by('detallemodeloevaluativo__orden')
#             col_nums = 40
#             for modelo in qsmodelo:
#                 hojadestino.write(fila, col_nums, modelo.valor if modelo else '', fuentenormal)
#                 col_nums += 1
#             fila += 1
#
#         libdestino.save(output_folder + libre_origen)
#         print(output_folder + libre_origen)
#         print("Proceso finalizado. . .")
#     except Exception as ex:
#         msg = ex.__str__()
#
#         textoerror = '{} Linea:{}'.format(str(ex), sys.exc_info()[-1].tb_lineno)
#         print(textoerror)
#         print(msg)
#
#
# def idusermoodlearreglocagada():
#     print(f'Desactivando matricula.....')
#     listaid = [45197,
#                37325,
#                33504,
#                37608,
#                40159,
#                28828,
#                30761,
#                30762,
#                28768,
#                31238,
#                27177,
#                25957,
#                18205,
#                18998,
#                14383,
#                27973,
#                6451,
#                10634,
#                7540,
#                18807,
#                5148,
#                20981,
#                6534,
#                17124,
#                26677,
#                27598,
#                10837,
#                22160,
#                21250,
#                21286,
#                45249,
#                6374,
#                2307,
#                116,
#                45258,
#                27733,
#                26721,
#                27729,
#                2469,
#                27205,
#                15118,
#                2788,
#                40362,
#                37870,
#                2817,
#                25065,
#                20474,
#                33778,
#                11206,
#                24094,
#                37291,
#                16732,
#                8008,
#                20494,
#                5285,
#                45368,
#                5373,
#                24477,
#                24728,
#                45706,
#                32050,
#                45362,
#                37263,
#                15302,
#                26710,
#                11564,
#                13224,
#                43163,
#                6569,
#                35498,
#                9658,
#                20247,
#                16983,
#                38585,
#                28221,
#                23590,
#                2943,
#                7662,
#                23314,
#                31213,
#                9128,
#                45453,
#                28186,
#                1972,
#                29237,
#                45672,
#                7658,
#                11052,
#                23119,
#                7657,
#                4374,
#                3778,
#                7668,
#                7671,
#                38595,
#                28188,
#                24096,
#                6627,
#                20672,
#                28678,
#                45557,
#                21450,
#                29540,
#                14236,
#                29517,
#                21731,
#                30136,
#                17043,
#                33090,
#                16762,
#                29820,
#                10679,
#                45330,
#                15354,
#                27762,
#                7032,
#                6384,
#                13673,
#                10457,
#                29651,
#                2941,
#                19133,
#                28232,
#                10759,
#                16105,
#                45061,
#                25416,
#                45689,
#                31511,
#                10860,
#                20552,
#                31320,
#                36606,
#                18101,
#                34083,
#                3723,
#                35672,
#                45271,
#                22837,
#                15596,
#                23844,
#                45906,
#                16200,
#                31178,
#                6219,
#                24968,
#                11912,
#                15925,
#                19483,
#                45624,
#                45285,
#                25367,
#                26711,
#                15964,
#                33620,
#                33801,
#                22515,
#                45290,
#                22925,
#                27094,
#                30632,
#                23196,
#                37179,
#                24454,
#                45374,
#                4491,
#                10671,
#                44682,
#                11592,
#                45491,
#                37993,
#                5649,
#                28983,
#                37972,
#                20438,
#                45277,
#                24884,
#                11441,
#                26115,
#                3339,
#                45287,
#                37055,
#                7888,
#                20289,
#                45612,
#                2527,
#                25353,
#                11099,
#                10553,
#                45295,
#                12028,
#                45741,
#                49809,
#                46188,
#                15026,
#                19631,
#                23072,
#                12054,
#                13809,
#                27814,
#                45962,
#                33577,
#                39985,
#                28775,
#                45394,
#                14667,
#                45670,
#                31807,
#                12387,
#                39988,
#                12344,
#                45802,
#                24517,
#                20930,
#                43343,
#                40404,
#                45334,
#                12625,
#                24196,
#                16556,
#                43613,
#                11696,
#                4749,
#                28930,
#                12636,
#                41563,
#                5938,
#                11457,
#                35035,
#                13806,
#                45273,
#                19115,
#                19705,
#                8347,
#                27162,
#                12765,
#                35904,
#                2069,
#                45613,
#                45532,
#                20317,
#                1029,
#                6301,
#                41736,
#                28834,
#                42873,
#                15887,
#                34972,
#                6813,
#                25969,
#                26535,
#                45825,
#                42422,
#                11576,
#                8031,
#                5072,
#                6831,
#                3388,
#                44681,
#                6077,
#                21052,
#                24165,
#                32381,
#                13936,
#                3641,
#                45615,
#                5080,
#                3895,
#                19238,
#                13703,
#                1503,
#                12151,
#                17169,
#                230,
#                5170,
#                18642,
#                42251,
#                34833,
#                26368,
#                2390,
#                22336,
#                13603,
#                7664,
#                4059,
#                25569,
#                45329,
#                13829,
#                13871,
#                7700,
#                7698,
#                27209,
#                12516,
#                13808,
#                31885,
#                26712,
#                3476,
#                26893,
#                28357,
#                26951,
#                13814,
#                40558,
#                14136,
#                14196,
#                44533,
#                14257,
#                28283,
#                45327,
#                13818,
#                18765,
#                8011,
#                46030,
#                13080,
#                45385,
#                26946,
#                25480,
#                27245,
#                34226,
#                30134,
#                16652,
#                29756,
#                29080,
#                45403,
#                5931,
#                6647,
#                15662,
#                8266,
#                37068,
#                45549,
#                27246,
#                45756,
#                25710,
#                6679,
#                16383,
#                16390,
#                4208,
#                23769,
#                16364,
#                20621,
#                16755,
#                45845,
#                16921,
#                45667,
#                45766,
#                46093,
#                29531,
#                45272,
#                45684,
#                19359,
#                26887,
#                26748,
#                1255,
#                36752,
#                42312,
#                29455]
#     countmatri = len(listaid)
#     cnmoodle = connections['moodle_db'].cursor()
#     count = 1
#     for l in listaid:
#         sql = """Select id, username From mooc_user Where id=%s""" % (l)
#         cnmoodle.execute(sql)
#         registro = cnmoodle.fetchall()
#         idusuario = registro[0][0]
#         username = registro[0][1]
#         persona__ = Persona.objects.filter(status=True, usuario__username=username)
#         if persona__.exists():
#             persona_ = persona__.first()
#             matriculas_ = Matricula.objects.filter(status=True, nivel__periodo__id=153, inscripcion__persona__usuario__username=username).first()
#             if matriculas_:
#                 materiasasignadas_ = matriculas_.materiaasignada_set.filter(status=True).exclude(materia__asignaturamalla__malla__carrera__id=34)
#                 print(f'{count}/{countmatri} {matriculas_.__str__()}')
#                 for materia in materiasasignadas_:
#                     materia.materia.crear_actualizar_un_estudiante_curso(moodle, 1, matriculas_)
#                     print(f'------------------Materia: {materia}')
#         count += 1
#
#
# def bloqeomatricula():
#     print(f'Desactivando matricula.....')
#     matriculas = Matricula.objects.filter(id__in=[551416,
#                                                   560574,
#                                                   557902,
#                                                   557902,
#                                                   557902,
#                                                   559691,
#                                                   549854,
#                                                   549854,
#                                                   541375,
#                                                   555082,
#                                                   555082,
#                                                   524014,
#                                                   524014,
#                                                   559999,
#                                                   553779,
#                                                   553779,
#                                                   542963,
#                                                   542963,
#                                                   542963,
#                                                   540313,
#                                                   560576,
#                                                   560576,
#                                                   528959,
#                                                   546918,
#                                                   546918,
#                                                   550000,
#                                                   558590,
#                                                   560877,
#                                                   560877,
#                                                   536845,
#                                                   536845,
#                                                   553882,
#                                                   554403,
#                                                   554403,
#                                                   536392,
#                                                   555016,
#                                                   545234,
#                                                   545234,
#                                                   556298,
#                                                   545256,
#                                                   545256,
#                                                   541783,
#                                                   537387,
#                                                   537387,
#                                                   548628,
#                                                   531138,
#                                                   531138,
#                                                   531009,
#                                                   549432,
#                                                   537709,
#                                                   559119,
#                                                   550297,
#                                                   531835,
#                                                   559913,
#                                                   529459,
#                                                   555913,
#                                                   555913,
#                                                   559159,
#                                                   555812,
#                                                   555812,
#                                                   537003,
#                                                   554751,
#                                                   554751,
#                                                   559005,
#                                                   559005,
#                                                   522074,
#                                                   522074,
#                                                   522074,
#                                                   522664,
#                                                   544372,
#                                                   559151,
#                                                   559151,
#                                                   558346,
#                                                   554218,
#                                                   549224,
#                                                   534952,
#                                                   534952,
#                                                   553241,
#                                                   553241,
#                                                   557711,
#                                                   541563,
#                                                   541563,
#                                                   541563,
#                                                   554358,
#                                                   559511,
#                                                   559511,
#                                                   522452,
#                                                   536127,
#                                                   536127,
#                                                   558442,
#                                                   558442,
#                                                   547423,
#                                                   518221,
#                                                   518221,
#                                                   544217,
#                                                   532975,
#                                                   532975,
#                                                   551086,
#                                                   551086,
#                                                   539575,
#                                                   552214,
#                                                   558815,
#                                                   546222,
#                                                   546222,
#                                                   537400,
#                                                   537400,
#                                                   559819,
#                                                   559819,
#                                                   555619,
#                                                   549693,
#                                                   549693,
#                                                   533029,
#                                                   533029,
#                                                   554846,
#                                                   560759,
#                                                   537570,
#                                                   558718,
#                                                   527937,
#                                                   537134,
#                                                   558356,
#                                                   558356,
#                                                   551480,
#                                                   551480,
#                                                   537622,
#                                                   544697,
#                                                   557861,
#                                                   548190,
#                                                   550598,
#                                                   535015,
#                                                   553453,
#                                                   543076,
#                                                   543076,
#                                                   536659,
#                                                   536634,
#                                                   560717,
#                                                   558319,
#                                                   558319,
#                                                   560206,
#                                                   560206,
#                                                   560206,
#                                                   525366,
#                                                   525366,
#                                                   525366,
#                                                   525366,
#                                                   553335,
#                                                   558940,
#                                                   558940,
#                                                   545505,
#                                                   536797,
#                                                   539374,
#                                                   539374,
#                                                   543137,
#                                                   543137,
#                                                   540247,
#                                                   540247,
#                                                   540564,
#                                                   540564,
#                                                   523989,
#                                                   536273,
#                                                   536273,
#                                                   536443,
#                                                   536443,
#                                                   551899,
#                                                   551899,
#                                                   552011,
#                                                   536541,
#                                                   536541,
#                                                   549691,
#                                                   549691,
#                                                   558549,
#                                                   534671,
#                                                   526921,
#                                                   559357,
#                                                   556114,
#                                                   533262,
#                                                   533961,
#                                                   545387,
#                                                   537449,
#                                                   555364,
#                                                   555364,
#                                                   546021,
#                                                   546021,
#                                                   554901,
#                                                   554901,
#                                                   525458,
#                                                   527826,
#                                                   559698,
#                                                   559698,
#                                                   542078,
#                                                   542078,
#                                                   551634,
#                                                   558341,
#                                                   535676,
#                                                   555730,
#                                                   545307,
#                                                   555541,
#                                                   549125,
#                                                   548925,
#                                                   559091,
#                                                   517297,
#                                                   517297,
#                                                   537337,
#                                                   537337,
#                                                   545980,
#                                                   545980,
#                                                   545936,
#                                                   550862,
#                                                   550862,
#                                                   539938,
#                                                   539938,
#                                                   551029,
#                                                   551029,
#                                                   542466,
#                                                   555565,
#                                                   555565,
#                                                   543641,
#                                                   543641,
#                                                   547523,
#                                                   532145,
#                                                   554810,
#                                                   546651,
#                                                   559297,
#                                                   559297,
#                                                   559297,
#                                                   543464,
#                                                   543464,
#                                                   536885,
#                                                   536885,
#                                                   554718,
#                                                   554718,
#                                                   559529,
#                                                   559529,
#                                                   559529,
#                                                   560280,
#                                                   560280,
#                                                   542762,
#                                                   545516,
#                                                   523601,
#                                                   523601,
#                                                   549736,
#                                                   557466,
#                                                   557466,
#                                                   554013,
#                                                   557760,
#                                                   540948,
#                                                   540948,
#                                                   539955,
#                                                   539955,
#                                                   543014,
#                                                   543014,
#                                                   558466,
#                                                   546269,
#                                                   546269,
#                                                   545724,
#                                                   545724,
#                                                   555254,
#                                                   555254,
#                                                   533385,
#                                                   533385,
#                                                   544761,
#                                                   544761,
#                                                   538109,
#                                                   538109,
#                                                   538983,
#                                                   548977,
#                                                   541666,
#                                                   541666,
#                                                   551647,
#                                                   532023,
#                                                   533729,
#                                                   519710,
#                                                   519710,
#                                                   545339,
#                                                   536447,
#                                                   558288,
#                                                   558288,
#                                                   558389,
#                                                   558389,
#                                                   548806,
#                                                   548806,
#                                                   517412,
#                                                   517412,
#                                                   541788,
#                                                   541788,
#                                                   542117,
#                                                   559258,
#                                                   532713,
#                                                   532713,
#                                                   532935,
#                                                   532935,
#                                                   532935,
#                                                   515915,
#                                                   515915,
#                                                   547813,
#                                                   547813,
#                                                   554801,
#                                                   554801,
#                                                   527293,
#                                                   527293,
#                                                   559373,
#                                                   555993,
#                                                   548205,
#                                                   548205,
#                                                   554050,
#                                                   554050,
#                                                   546215,
#                                                   546215,
#                                                   542235,
#                                                   542235,
#                                                   560366,
#                                                   560366,
#                                                   543837,
#                                                   543837,
#                                                   540902,
#                                                   542281,
#                                                   542281,
#                                                   557434,
#                                                   557434,
#                                                   557434,
#                                                   523629,
#                                                   523629,
#                                                   527506,
#                                                   527506,
#                                                   559969,
#                                                   532953,
#                                                   532953,
#                                                   537080,
#                                                   537080,
#                                                   536942,
#                                                   536942,
#                                                   557888,
#                                                   560723,
#                                                   539193,
#                                                   539193,
#                                                   550059,
#                                                   534114,
#                                                   534114,
#                                                   548028,
#                                                   548028,
#                                                   539848,
#                                                   539848,
#                                                   536758,
#                                                   536758,
#                                                   536758,
#                                                   526664,
#                                                   536016,
#                                                   547207,
#                                                   558432,
#                                                   559715,
#                                                   540433,
#                                                   540433,
#                                                   526754,
#                                                   526754,
#                                                   557359,
#                                                   557359,
#                                                   542181,
#                                                   542181,
#                                                   549685,
#                                                   549685,
#                                                   558135,
#                                                   554779,
#                                                   533495,
#                                                   533495,
#                                                   542867,
#                                                   542867,
#                                                   539472,
#                                                   536929,
#                                                   559300,
#                                                   542585,
#                                                   533826,
#                                                   533826,
#                                                   542839,
#                                                   557577,
#                                                   533973,
#                                                   533973,
#                                                   557718,
#                                                   557718,
#                                                   541259,
#                                                   531818,
#                                                   546343,
#                                                   546343,
#                                                   546343,
#                                                   554723,
#                                                   524474,
#                                                   524474,
#                                                   560569,
#                                                   560569,
#                                                   560569,
#                                                   560834,
#                                                   553924,
#                                                   553924,
#                                                   536918,
#                                                   536918,
#                                                   555198,
#                                                   558476,
#                                                   523880,
#                                                   550371,
#                                                   558308,
#                                                   558773,
#                                                   560573,
#                                                   560573,
#                                                   532755,
#                                                   532755,
#                                                   559279,
#                                                   534301,
#                                                   541170,
#                                                   541170,
#                                                   520758,
#                                                   520758,
#                                                   523081,
#                                                   560200,
#                                                   541501,
#                                                   541501,
#                                                   614515,
#                                                   560714,
#                                                   560714,
#                                                   560714,
#                                                   533461,
#                                                   558126,
#                                                   559771,
#                                                   559771,
#                                                   545398,
#                                                   545398,
#                                                   526105,
#                                                   526105,
#                                                   559930,
#                                                   559930,
#                                                   551103,
#                                                   548787,
#                                                   548787,
#                                                   527222,
#                                                   527222,
#                                                   543342,
#                                                   541680,
#                                                   535025,
#                                                   535025,
#                                                   539660,
#                                                   539660,
#                                                   545043,
#                                                   545043,
#                                                   559273,
#                                                   559056,
#                                                   521969,
#                                                   521969,
#                                                   538882,
#                                                   557945,
#                                                   543528,
#                                                   543528,
#                                                   537696,
#                                                   537696,
#                                                   515300,
#                                                   515300,
#                                                   550044,
#                                                   550044,
#                                                   560471,
#                                                   560471,
#                                                   560471,
#                                                   521835,
#                                                   523472,
#                                                   559230,
#                                                   559230,
#                                                   553074,
#                                                   553074,
#                                                   547645,
#                                                   547645,
#                                                   560068,
#                                                   560068,
#                                                   560068,
#                                                   541628,
#                                                   520139,
#                                                   520139,
#                                                   555458,
#                                                   555458,
#                                                   557391,
#                                                   554799,
#                                                   541651,
#                                                   556496,
#                                                   560116,
#                                                   560116,
#                                                   552923,
#                                                   559401,
#                                                   559401,
#                                                   545421,
#                                                   545421,
#                                                   559794,
#                                                   559794,
#                                                   519893,
#                                                   519893,
#                                                   560290,
#                                                   531868,
#                                                   531868,
#                                                   521834,
#                                                   558472,
#                                                   560163,
#                                                   560163,
#                                                   541009,
#                                                   560170,
#                                                   560170,
#                                                   529607,
#                                                   529607,
#                                                   533172,
#                                                   549120,
#                                                   545446,
#                                                   550396,
#                                                   560686,
#                                                   560686,
#                                                   536883,
#                                                   560214,
#                                                   518703,
#                                                   518703,
#                                                   539846,
#                                                   539846,
#                                                   558048,
#                                                   523050,
#                                                   542944,
#                                                   542944,
#                                                   542944,
#                                                   522693,
#                                                   558806,
#                                                   546085,
#                                                   546085,
#                                                   544083,
#                                                   544083,
#                                                   523871,
#                                                   523871,
#                                                   555130,
#                                                   555130,
#                                                   541486,
#                                                   541486,
#                                                   542778,
#                                                   541778,
#                                                   543113,
#                                                   543113,
#                                                   535898,
#                                                   535898,
#                                                   530207,
#                                                   530207,
#                                                   543236,
#                                                   543236,
#                                                   549065,
#                                                   549065,
#                                                   554941,
#                                                   527259,
#                                                   555148,
#                                                   555148,
#                                                   545799,
#                                                   545799,
#                                                   552319,
#                                                   552319,
#                                                   525886,
#                                                   525886,
#                                                   540904,
#                                                   546055,
#                                                   546055])
#     countmatri = len(matriculas)
#     cnmoodle = connections['moodle_db'].cursor()
#     count = 1
#     for m in matriculas:
#         # idusermoodle = m.inscripcion.persona.idusermoodle
#         username = m.inscripcion.persona.usuario.username
#         print(f'{count}/{countmatri} {m.__str__()}')
#         m.bloqueomatricula = True
#         m.save()
#         if username:
#             sql = f"Select id, username From mooc_user Where username='{username}'"
#             cnmoodle.execute(sql)
#             registro = cnmoodle.fetchall()
#             # idusuario = registro[0][0]
#             try:
#                 username = registro[0][1]
#                 # Asignar estado suspended = 1 para que no pueda acceder al aula virtual
#                 sql = f"Update mooc_user Set suspended=1 Where username='{username}'"
#                 cnmoodle.execute(sql)
#             except Exception as ex:
#                 print(f'********--------{username}--------********')
#         count += 1
#
#
# def desbloqeomatricula():
#     print(f'Activando matricula.....')
#
#     matriculas = Matricula.objects.filter(id__in=[
#         # 539193,
#                                                   # 549224,
#                                                   # 522693,
#                                                   # 545256,
#                                                   # 541009,
#                                                   # 543342,
#                                                   # 559698,
#                                                   # 529459,
#                                                   # 544217,
#                                                   # 555198,
#                                                   # 559715,
#                                                   # 545505,
#                                                   # 560280,
#                                                   # 558476,
#                                                   # 551634,
#                                                   # 526754,
#                                                   # 531818,
#                                                   # 536127,
#                                                   # 524014,
#                                                   # 550371,
#                                                   # 532023,
#                                                   # 521835,
#                                                   # 559819,
#                                                   # 546215,
#                                                   # 541259,
#                                                   # 539846,
#                                                   # 554801,
#                                                   # 542235,
#                                                   # 544697,
#                                                   # 558319,
#                                                   # 549693,
#                                                   # 540904,
#                                                   # 534671,
#                                                   # 547645,
#                                                   # 543641,
#                                                   # 540313,
#                                                   # 551029,
#                                                   # 552214,
#                                                   # 544761,
#                                                   # 555619,
#                                                   # 549736,
#                                                   # 552319,
#                                                   # 539848,
#                                                   # 530207,
#                                                   # 545043,
#                                                   # 560068,
#                                                   # 554358,
#                                                   # 548028,
#                                                   # 558389,
#                                                   # 560714,
#                                                   # 560717,
#                                                   # 542078,
#                                                   # 555082,
#                                                   # 552923,
#                                                   # 535015,
#                                                   # 557888,
#                                                   # 525458,
#                                                   # 549120,
#                                                   # 527259,
#                                                   # 531835,
#                                                   # 523081,
#                                                   # 555993,
#                                                   # 556114,
#                                                   # 557466,
#                                                   # 539472,
#                                                   # 539575,
#                                                   554810,
#                                                   557718,
#                                                   555812,
#                                                   # 546021,
#                                                   # 560200,
#                                                   # 554723,
#                                                   ])
#     countmatri = len(matriculas)
#     cnmoodle = connections['moodle_db'].cursor()
#     count = 1
#     for m in matriculas:
#         username = m.inscripcion.persona.usuario.username
#         print(f'{count}/{countmatri} {m.__str__()} - {m.id}')
#         m.bloqueomatricula = False
#         m.retiradomatricula = False
#         m.save()
#         materia_ = m.materiaasignada_set.filter(status=True).update(retiramateria=False)
#         if username:
#             sql = f"Select id, username From mooc_user Where username='{username}'"
#             cnmoodle.execute(sql)
#             registro = cnmoodle.fetchall()
#             # idusuario = registro[0][0]
#             username = registro[0][1]
#             # Asignar estado suspended = 1 para que no pueda acceder al aula virtual
#             sql = f"Update mooc_user Set suspended=0 Where username='{username}'"
#             cnmoodle.execute(sql)
#
#

# def subir_bitacora_personal_22():
#     archivo_ = 'actividades_llerena1'
#     # archivo_ = 'PLANTILLA_INSTITUCION_MJRV_CNE ELECCIONES_SECCIONALES_CPCCS_2023'
#     # url_archivo = "{}/{}.xlsx".format(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), archivo_)
#     url_archivo = "{}/media/{}.xlsx".format(SITE_STORAGE, archivo_)
#     wb = openpyxl.load_workbook(filename=url_archivo)
#     ws = wb.get_sheet_by_name("2023 1s")
#     cadena = ''
#     linea, excluidos, conexito = 0, 0, 0
#     worksheet = ws
#     lis_excluidos = []
#     print('Iniciando....')
#     linea_archivo = 1
#     col_fecha = 0
#     col_hora = 1
#     col_titulo = 2
#     col_actividad = 3
#     col_py_afectados = 4
#     departamento = Departamento.objects.get(pk=93)
#     persona = Persona.objects.get(pk=28164)
#     for row in worksheet.iter_rows(min_row=0):
#         if linea > 1:
#             currentValues, cadena = [], ''
#             for cell in row:
#                 cadena += str(cell.value) + ' '
#                 currentValues.append(str(cell.value))
#             print(currentValues)
#             f = convertir_fecha_hora_invertida(f'{currentValues[0].split()[0]} {currentValues[1]}')
#             with transaction.atomic():
#                 try:
#                     if not BitacoraActividadDiaria.objects.filter(fecha=f).exists():
#                         eBitacoraActividadDiaria = BitacoraActividadDiaria(titulo=currentValues[2],
#                                                                            departamento=departamento,
#                                                                            fecha=f,
#                                                                            persona=persona,
#                                                                            descripcion=f"{currentValues[3]} en los archivos: {currentValues[4]}",
#                                                                            tiposistema=2)
#                         eBitacoraActividadDiaria.save(usuario_id=persona.usuario.id)
#                         print(f"Se guardo registro de fecha {eBitacoraActividadDiaria.fecha.__str__()}")
#                 except Exception as ex:
#                     transaction.set_rollback(True)
#                     print(f"No se guardo registro de fecha {f.__str__()}")
#         linea += 1
#
# subir_bitacora_personal_22()
#
#
# def estudianteseptimo():
#     cadena = ''
#     linea, excluidos, conexito = 0, 0, 0
#     try:
#         archivo_ = 'ESTUDIANTES_7MO'
#         # archivo_ = 'PLANTILLA_INSTITUCION_MJRV_CNE ELECCIONES_SECCIONALES_CPCCS_2023'
#         # url_archivo = "{}/{}.xlsx".format(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), archivo_)
#         url_archivo = "{}/media/{}.xlsx".format(SITE_STORAGE, archivo_)
#         print('Leyendo....')
#         # wb = openpyxl.load_workbook(filename="{}/media/{}.xlsx".format(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), archivo_))
#         wb = openpyxl.load_workbook(filename=url_archivo)
#         ws = wb.get_sheet_by_name("Hoja1")
#         # worksheet = wb["Listado"]
#         # inscrip = Inscripcion.objects.get(id=81426)
#         # vali = inscrip.todas_materias_aprobadas_rango_nivel2(1, 7)
#         # print(vali)
#         worksheet = ws
#         lis_excluidos = []
#         print('Iniciando....')
#         linea_archivo = 1
#         for row in worksheet.iter_rows(min_row=0):
#             if linea >= 1:
#                 currentValues, cadena = [], ''
#                 worksheet["K{}".format(linea_archivo)].value = ''
#                 worksheet["L{}".format(linea_archivo)].value = ''
#                 for cell in row:
#                     cadena += str(cell.value) + ' '
#                     currentValues.append(str(cell.value))
#                 apellido1_, apellido2_, nombre_ = currentValues[5], currentValues[6], currentValues[7]
#                 nombrecompleto = f'{apellido1_} {apellido2_} {nombre_}'
#                 matriculaid = currentValues[9]
#                 matricula_ = Matricula.objects.get(id=matriculaid)
#                 inscripcion_ = matricula_.inscripcion
#                 worksheet["K{}".format(linea_archivo)].value = inscripcion_.todas_materias_aprobadas_rango_nivel2(1, 7)
#                 worksheet["L{}".format(linea_archivo)].value = inscripcion_.todas_materias_aprobadas_rango_nivel2(1, 6)
#                 conexito += 1
#                 print("Linea {}/? - Persona: {}".format(linea, nombrecompleto))
#             linea += 1
#             linea_archivo += 1
#             if linea in (25, 50, 400, 800, 1200, 3000, 5000, 6000, 8000, 10000, 15000, 18000, 20000, 20500, 30000, 30500):
#                 wb.save(url_archivo)
#                 print("Guardado Rapido Linea {} . .".format(linea))
#         print('Total Leidos con exito: {}'.format(conexito))
#         print('Total Excluidos: {}'.format(excluidos))
#         print(lis_excluidos)
#         wb.save(url_archivo)
#         # enviar_mensaje_bot_telegram('FINALIZO CON EXITO: {}'.format(url_archivo))
#     except Exception as ex:
#         textoerror = '{} Linea:{} - Info: {} / Lectura: {}'.format(str(ex), sys.exc_info()[-1].tb_lineno, cadena, linea)
#         print(textoerror)
#         # enviar_mensaje_bot_telegram(textoerror)
#
#
# def matrizasignaturamateria():
#     cadena = ''
#     linea, excluidos, conexito = 0, 0, 0
#     try:
#         archivo_ = 'listado_materias'
#         # archivo_ = 'PLANTILLA_INSTITUCION_MJRV_CNE ELECCIONES_SECCIONALES_CPCCS_2023'
#         # url_archivo = "{}/{}.xlsx".format(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), archivo_)
#         url_archivo = "{}/media/{}.xlsx".format(SITE_STORAGE, archivo_)
#         print('Leyendo....')
#         # wb = openpyxl.load_workbook(filename="{}/media/{}.xlsx".format(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), archivo_))
#         wb = openpyxl.load_workbook(filename=url_archivo)
#         ws = wb.get_sheet_by_name("Hoja1")
#         # worksheet = wb["Listado"]
#         worksheet = ws
#         lis_excluidos = []
#         print('Iniciando....')
#         worksheet["A{}".format(1)].value = 'COORDINACION'
#         worksheet["B{}".format(1)].value = 'CARRERA'
#         worksheet["C{}".format(1)].value = 'MATERIA'
#         worksheet["D{}".format(1)].value = 'NIVEL'
#         worksheet["E{}".format(1)].value = 'PARALELO'
#         worksheet["F{}".format(1)].value = 'HORAS'
#         worksheet["G{}".format(1)].value = 'PROFESOR PRINCIPAL'
#         worksheet["H{}".format(1)].value = 'CAMPO ESPECIFICO'
#         worksheet["I{}".format(1)].value = 'CAMPO AMPLIO'
#         worksheet["J{}".format(1)].value = 'CAMPO DETALLADO'
#         linea_archivo = 2
#         listmaterias_ = Materia.objects.filter(status=True, nivel__periodo__id=153).order_by('asignaturamalla__malla__carrera__alias')
#         for mat in listmaterias_:
#             worksheet["A{}".format(linea_archivo)].value = mat.asignaturamalla.malla.carrera.mi_coordinacion().__str__()
#             worksheet["B{}".format(linea_archivo)].value = mat.asignaturamalla.malla.carrera.__str__()
#             worksheet["C{}".format(linea_archivo)].value = mat.asignatura.nombre
#             worksheet["D{}".format(linea_archivo)].value = mat.asignaturamalla.nivelmalla.__str__()
#             worksheet["E{}".format(linea_archivo)].value = mat.paralelo.__str__()
#             worksheet["F{}".format(linea_archivo)].value = mat.horas
#             worksheet["G{}".format(linea_archivo)].value = mat.profesor_principal().persona.__str__() if mat.profesor_principal() else ''
#             campoespecifico, campoamplio, campodetallado = '', '', ''
#
#             campoespecificofirst_ = mat.asignaturamalla.areaconocimientotitulacion.__str__() if mat.asignaturamalla.areaconocimientotitulacion else ''
#             campoampliofirst_ = mat.asignaturamalla.subareaconocimiento.__str__() if mat.asignaturamalla.subareaconocimiento else ''
#             campodetalladofirst_ = mat.asignaturamalla.subareaespecificaconocimiento.__str__() if mat.asignaturamalla.subareaespecificaconocimiento else ''
#             campoespecifico += f'{campoespecificofirst_}'
#             campoamplio += f'{campoampliofirst_}'
#             campodetallado += f'{campodetalladofirst_}'
#             worksheet["H{}".format(linea_archivo)].value = campoespecifico
#             worksheet["I{}".format(linea_archivo)].value = campoamplio
#             worksheet["J{}".format(linea_archivo)].value = campodetallado
#             print("Linea {}/{} - {}\nEspecifico: {}\nAmplio: {}\nDetallado: {}".format(linea, len(listmaterias_), mat.__str__(), campoespecifico, campoamplio, campodetallado))
#             linea += 1
#             linea_archivo += 1
#             if linea in (25, 50, 400, 800, 1200, 3000, 5000, 6000, 8000, 10000, 15000, 18000, 20000, 20500, 30000, 30500):
#                 wb.save(url_archivo)
#                 print("Guardado Rapido Linea {} . .".format(linea))
#         print('Total Leidos con exito: {}'.format(conexito))
#         print('Total Excluidos: {}'.format(excluidos))
#         print(lis_excluidos)
#         wb.save(url_archivo)
#         # enviar_mensaje_bot_telegram('FINALIZO CON EXITO: {}'.format(url_archivo))
#     except Exception as ex:
#         textoerror = '{} Linea:{} - Info: {} / Lectura: {}'.format(str(ex), sys.exc_info()[-1].tb_lineno, cadena, linea)
#         print(textoerror)
#         # enviar_mensaje_bot_telegram(textoerror)
#
#
# def matrizmesa():
#     cadena = ''
#     linea, excluidos, conexito = 0, 0, 0
#     try:
#         archivo_ = 'LISTADO_MESA'
#         # archivo_ = 'PLANTILLA_INSTITUCION_MJRV_CNE ELECCIONES_SECCIONALES_CPCCS_2023'
#         # url_archivo = "{}/{}.xlsx".format(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), archivo_)
#         url_archivo = "{}/media/{}.xlsx".format(SITE_STORAGE, archivo_)
#         print('Leyendo....')
#         # wb = openpyxl.load_workbook(filename="{}/media/{}.xlsx".format(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), archivo_))
#         wb = openpyxl.load_workbook(filename=url_archivo)
#         ws = wb.get_sheet_by_name("UNEMI GENERAL")
#         # worksheet = wb["Listado"]
#         worksheet = ws
#         lis_excluidos = []
#         print('Iniciando....')
#         linea_archivo = 1
#         for row in worksheet.iter_rows(min_row=0):
#             if linea >= 3:
#                 currentValues, cadena = [], ''
#                 worksheet["G{}".format(linea_archivo)].value = ''
#                 worksheet["H{}".format(linea_archivo)].value = ''
#                 for cell in row:
#                     cadena += str(cell.value) + ' '
#                     currentValues.append(str(cell.value))
#                 nombres = currentValues[1]
#                 lista_nombres = nombres.split()
#                 if len(lista_nombres) > 3:
#                     apellido1_, apellido2_, nombre_ = lista_nombres[0], lista_nombres[1], f'{lista_nombres[2]} {lista_nombres[3]}'
#                 else:
#                     apellido1_, apellido2_, nombre_ = lista_nombres[0], lista_nombres[1], f'{lista_nombres[2]}'
#                 qsbasepersona = Persona.objects.filter(status=True, apellido1=apellido1_, apellido2=apellido2_, nombres__icontains=nombre_)
#                 if qsbasepersona.exists():
#                     persona_ = qsbasepersona.first()
#                     worksheet["G{}".format(linea_archivo)].value = persona_.email
#                     worksheet["H{}".format(linea_archivo)].value = persona_.emailinst
#                     conexito += 1
#                 else:
#                     lis_excluidos.append({'cedula': nombres, 'obs': 'No existe persona'})
#                     excluidos += 1
#                 print("Linea {}/? - Persona: {}".format(linea, nombres))
#             linea += 1
#             linea_archivo += 1
#             if linea in (25, 50, 400, 800, 1200, 3000, 5000, 6000, 8000, 10000, 15000, 18000, 20000, 20500, 30000, 30500):
#                 wb.save(url_archivo)
#                 print("Guardado Rapido Linea {} . .".format(linea))
#         print('Total Leidos con exito: {}'.format(conexito))
#         print('Total Excluidos: {}'.format(excluidos))
#         print(lis_excluidos)
#         wb.save(url_archivo)
#         # enviar_mensaje_bot_telegram('FINALIZO CON EXITO: {}'.format(url_archivo))
#     except Exception as ex:
#         textoerror = '{} Linea:{} - Info: {} / Lectura: {}'.format(str(ex), sys.exc_info()[-1].tb_lineno, cadena, linea)
#         print(textoerror)
#         # enviar_mensaje_bot_telegram(textoerror)
#
#
# def cargar_modfavoritos_1():
#     # ids = [472]
#     ids = [472]
#     modulos_ = Modulo.objects.filter(id__in=ids)
#     docentes_ = PerfilUsuario.objects.filter(status=True, visible=True, administrativo__isnull=False)
#     for d_ in docentes_:
#         existe_ = 'NO'
#         qsfav = MenuFavoriteProfile.objects.filter(profile=d_, setting__id=1)
#         fav = None
#         if qsfav.exists():
#             existe_ = 'SI'
#             fav = qsfav.first()
#         else:
#             fav = MenuFavoriteProfile(profile=d_, setting_id=1)
#             fav.save()
#         for m in modulos_:
#             if not fav.modules.filter(id=m.id).exists():
#                 fav.modules.add(m)
#                 fav.save()
#                 print(m)
#         print(f'{d_.id} - {d_.persona.cedula} - {d_.persona} -- EXISTE: {existe_}')
#
#
# def cargar_modfavoritos():
#     asbase = MenuFavoriteProfile.objects.get(pk=38)
#     ids = [472]
#     modulos_ = Modulo.objects.filter(id__in=ids)
#     docentes_ = PerfilUsuario.objects.filter(status=True, profesor__isnull=False, visible=True)
#     for d_ in docentes_:
#         existe_ = 'NO'
#         qsfav = MenuFavoriteProfile.objects.filter(profile=d_, setting__id=2)
#         fav = None
#         if qsfav.exists():
#             existe_ = 'SI'
#             fav = qsfav.first()
#         else:
#             fav = MenuFavoriteProfile(profile=d_, setting_id=2)
#             fav.save()
#         for m in modulos_:
#             if not fav.modules.filter(id=m.id).exists():
#                 fav.modules.add(m)
#                 fav.save()
#                 print(m)
#         print(f'{d_.id} - {d_.persona.cedula} - {d_.persona} -- EXISTE: {existe_}')
#
#
# def arreglosalidas():
#     try:
#         salidas_ = SalidaProducto.objects.filter(numerodocumento__in=[3040, 3041, 3042, 3043, 3045, 3046, 3047, 3052, 3057, 3059, 3063, 3067, 3068, 3081, 3082, 3085, 3086, 3088, 3090, 3091, 3093, 3097, 3103, 3104, 3107, 3108, 3110, 3113, 3116, 3117], status=True)
#         for s in salidas_:
#             print('---------------------------------------')
#             print(f'TOTAL: {s.valor}')
#             for p in s.productos.filter(producto_id=102):
#                 print(f'ANTERIOR: --- {p.producto.id} {p}: Cant.{p.cantidad} ${p.costo} - ${p.valor}')
#                 p.costo = 2.70
#                 p.valor = float(p.cantidad) * 2.70
#                 p.save()
#                 p1 = DetalleSalidaProducto.objects.get(pk=p.id)
#                 print(f'ACTUAL: --- {p1.producto.id} {p}: Cant.{p1.cantidad} ${p1.costo} - ${p1.valor}')
#             print('-------------FIN---------------')
#             print(f'------ ANTERIOR: ${s.valor}')
#             s.valor = null_to_decimal(s.productos.aggregate(suma=Sum('valor'))['suma'])
#             s.save()
#             print(f"------ ACTUAL: ${null_to_decimal(s.productos.aggregate(suma=Sum('valor'))['suma'])}")
#     except Exception as ex:
#         print(ex)
#
#
# def webpushalerta():
#     pers = Persona.objects.get(pk=22213)
#     usernotify = User.objects.get(pk=pers.usuario.id)
#     noti = Notificacion(cuerpo='Reporte Listo', titulo='EJEMPLO PRUEBA',
#                         destinatario=pers, url="{}reportes/matrices/".format(MEDIA_URL), prioridad=1,
#                         app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
#                         tipo=2, en_proceso=False, error=False)
#     noti.save()
#     noti = Notificacion(cuerpo='Reporte Listo', titulo='EJEMPLO PRUEBA',
#                         destinatario=pers, url="{}reportes/matrices/".format(MEDIA_URL), prioridad=1,
#                         app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
#                         tipo=1, en_proceso=False, error=True)
#     noti.save()
#     noti = Notificacion(cuerpo='Reporte Listo', titulo='EJEMPLO PRUEBA',
#                         destinatario=pers, url="{}reportes/matrices/".format(MEDIA_URL), prioridad=1,
#                         app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
#                         tipo=2, en_proceso=False, error=True)
#     noti.save()
#     noti = Notificacion(cuerpo='Reporte Listo', titulo='EJEMPLO PRUEBA',
#                         destinatario=pers, url="{}reportes/matrices/".format(MEDIA_URL), prioridad=1,
#                         app_label='SGA', fecha_hora_visible=datetime.now() + timedelta(days=1),
#                         tipo=3, en_proceso=False, error=True)
#     noti.save()
#     send_user_notification(user=usernotify, payload={
#         "head": "Excel terminado",
#         "body": 'Prueba mi dog',
#         "action": "notificacion",
#         "timestamp": time.mktime(datetime.now().timetuple()),
#         "url": "{}reportes/matrices/".format(MEDIA_URL),
#         "noti_mensaje": 'Notificación Exitosa'
#     }, ttl=500)
#
#
# def traerhijosfolder(lista, nivel):
#     try:
#         for l in lista:
#             cadena_ = (str(nivel).zfill(nivel)).replace('0', '-')
#             print(f"---{cadena_} [{l.parent}] {l.nombre} ({l.cant_archivos()} Archivos - {l.cant_carpetas()} Carpetas)")
#             if l.traerhijas():
#                 traerhijosfolder(l.traerhijas(), nivel + 1)
#     except Exception as ex:
#         pass
#
#
# def traerarbol():
#     departamentos = DepartamentoArchivos.objects.filter(status=True)
#     for d in departamentos:
#         print(f"{d.departamento.nombre}")
#         for g in d.traergestiones():
#             print(f"-- {g.gestion.descripcion}")
#             if g.traerprimernivel():
#                 traerhijosfolder(g.traerprimernivel(), 1)
#
#
# def enviar_mensaje_bot_telegram(mensaje):
#     import requests
#     json_arr = []
#     try:
#         api = '1954045154:AAFK8CJo2Wr3nWpS-acBRgchZzcz2qxUKSU'
#         cgomez, clocke, rviteri = '900543432', '838621184', '2006141724'
#         chats = [rviteri]
#         for x in chats:
#             data = {'chat_id': x, 'text': mensaje, 'parse_mode': 'HTML'}
#             url = "https://api.telegram.org/bot{}/sendMessage".format(api)
#             json_arr.append(requests.post(url, data).json())
#     except Exception as ex:
#         print("TELEGRAM ERROR" + str(ex))
#     return json_arr
#
#
# def matrizcne():
#     cadena = ''
#     linea, excluidos, conexito = 0, 0, 0
#     try:
#         archivo_ = 'PLANTILLA_INSTITUCION_EDUCACION_SUPERIOR_MJRV_CNE_ELECCIONES__SECCIONALES_CPCCS_2023'
#         # archivo_ = 'PLANTILLA_INSTITUCION_MJRV_CNE ELECCIONES_SECCIONALES_CPCCS_2023'
#         # url_archivo = "{}/{}.xlsx".format(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), archivo_)
#         url_archivo = "{}/media/{}.xlsx".format(SITE_STORAGE, archivo_)
#         print('Leyendo....')
#         # wb = openpyxl.load_workbook(filename="{}/media/{}.xlsx".format(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), archivo_))
#         wb = openpyxl.load_workbook(filename=url_archivo)
#         ws = wb.get_sheet_by_name("LISTADO")
#         # worksheet = wb["Listado"]
#         worksheet = ws
#         lis_excluidos = []
#         print('Iniciando....')
#         linea_archivo = 1
#         for row in worksheet.iter_rows(min_row=0):
#             instruccion_ = 'Estudiante Universitario'
#             if linea >= 1:
#                 currentValues, cadena = [], ''
#                 worksheet["C{}".format(linea_archivo)].value = ''
#                 worksheet["D{}".format(linea_archivo)].value = ''
#                 worksheet["E{}".format(linea_archivo)].value = ''
#                 worksheet["F{}".format(linea_archivo)].value = ''
#                 worksheet["G{}".format(linea_archivo)].value = ''
#                 worksheet["I{}".format(linea_archivo)].value = ''
#                 worksheet["J{}".format(linea_archivo)].value = ''
#                 worksheet["K{}".format(linea_archivo)].value = ''
#                 worksheet["L{}".format(linea_archivo)].value = ''
#                 for cell in row:
#                     cadena += str(cell.value) + ' '
#                     currentValues.append(str(cell.value))
#                 cedula_ = currentValues[0]
#                 qsbasepersona = Persona.objects.filter(status=True, cedula__icontains=cedula_)
#                 if qsbasepersona.exists():
#                     persona_ = qsbasepersona.first()
#                     if persona_.titulacion_set.filter(status=True, titulo__nivel__nivel=3).exists():
#                         instruccion_ = 'Profesional'
#                     worksheet["C{}".format(linea_archivo)].value = instruccion_
#                     worksheet["D{}".format(linea_archivo)].value = persona_.direccion if persona_.direccion else ''
#                     worksheet["E{}".format(linea_archivo)].value = persona_.num_direccion if persona_.num_direccion else ''
#                     worksheet["F{}".format(linea_archivo)].value = persona_.direccion2 if persona_.direccion2 else ''
#                     worksheet["G{}".format(linea_archivo)].value = persona_.referencia if persona_.referencia else ''
#                     worksheet["I{}".format(linea_archivo)].value = persona_.parroquia.nombre if persona_.parroquia else ''
#                     worksheet["J{}".format(linea_archivo)].value = persona_.sector if persona_.sector else ''
#                     worksheet["K{}".format(linea_archivo)].value = persona_.sector if persona_.sector else ''
#                     worksheet["L{}".format(linea_archivo)].value = persona_.telefono_conv if persona_.telefono_conv else ''
#                     conexito += 1
#                 else:
#                     lis_excluidos.append({'cedula': cedula_, 'obs': 'No existe persona'})
#                     excluidos += 1
#                 print("Linea {}/? - Persona: {}".format(linea, cedula_))
#             linea += 1
#             linea_archivo += 1
#             if linea in (25, 50, 400, 800, 1200, 3000, 5000, 6000, 8000, 10000, 15000, 18000, 20000, 20500, 30000, 30500):
#                 wb.save(url_archivo)
#                 print("Guardado Rapido Linea {} . .".format(linea))
#         print('Total Leidos con exito: {}'.format(conexito))
#         print('Total Excluidos: {}'.format(excluidos))
#         # print(lis_excluidos)
#         wb.save(url_archivo)
#         enviar_mensaje_bot_telegram('FINALIZO CON EXITO: {}'.format(url_archivo))
#     except Exception as ex:
#         textoerror = '{} Linea:{} - Info: {} / Lectura: {}'.format(str(ex), sys.exc_info()[-1].tb_lineno, cadena, linea)
#         print(textoerror)
#         enviar_mensaje_bot_telegram(textoerror)
#
#
# def matrizadmision():
#     cadena = ''
#     linea, excluidos, conexito = 0, 0, 0
#     try:
#         archivo_ = 'ADM_UNEMI_2021_2'
#         # url_archivo = "{}/{}.xlsx".format(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), archivo_)
#         url_archivo = "{}/media/{}.xlsx".format(SITE_STORAGE, archivo_)
#         print('Leyendo....')
#         # wb = openpyxl.load_workbook(filename="{}/media/{}.xlsx".format(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), archivo_))
#         wb = openpyxl.load_workbook(filename=url_archivo)
#         ws = wb.get_sheet_by_name("Listado")
#         # worksheet = wb["Listado"]
#         worksheet = ws
#         lis_excluidos = []
#         print('Iniciando....')
#         linea_archivo = 1
#         for row in worksheet.iter_rows(min_row=0):
#             numinscripciones, nummatriculas = 0, 0
#             var1, var2, var3, var4, var5, var6, var7, var8 = 'NO', 'NO', 'NO', 'NO', 'NO', 'NO', 'NO', 'NO'
#             if linea >= 2:
#                 currentValues, cadena = [], ''
#                 worksheet["N{}".format(linea_archivo)].value = ''
#                 worksheet["O{}".format(linea_archivo)].value = ''
#                 worksheet["P{}".format(linea_archivo)].value = ''
#                 worksheet["Q{}".format(linea_archivo)].value = ''
#                 worksheet["R{}".format(linea_archivo)].value = ''
#                 worksheet["S{}".format(linea_archivo)].value = ''
#                 worksheet["T{}".format(linea_archivo)].value = ''
#                 worksheet["U{}".format(linea_archivo)].value = ''
#                 for cell in row:
#                     cadena += str(cell.value) + ' '
#                     currentValues.append(str(cell.value))
#                 cedula_ = currentValues[10]
#                 carrera_ = currentValues[2]
#                 qsbasepersona = Persona.objects.filter(status=True, cedula__icontains=cedula_)
#                 if qsbasepersona.exists():
#                     persona_ = qsbasepersona.first()
#                     qsbaseinscripcion = Inscripcion.objects.filter(status=True, persona=persona_, coordinacion_id=9, carrera__nombre__unaccent__icontains=carrera_.strip())
#                     numinscripciones = qsbaseinscripcion.count()
#                     if qsbaseinscripcion.exists():
#                         qsmatriculas = Matricula.objects.filter(status=True, inscripcion=qsbaseinscripcion.first(), automatriculaadmision=True).order_by('id')
#                         nummatriculas = len(qsmatriculas)
#                         if nummatriculas >= 1:
#                             var1 = 'SI' if nummatriculas >= 1 else 'NO'
#                             var2 = qsmatriculas.first().materias_aprobadas_todas()
#
#                             worksheet["N{}".format(linea_archivo)].value = var1
#                             worksheet["O{}".format(linea_archivo)].value = 'SI' if var2 else ''
#
#                             if nummatriculas >= 2:
#                                 var3 = 'SI' if nummatriculas >= 2 else ''
#                                 var4 = qsmatriculas[1].materias_aprobadas_todas()
#
#                                 worksheet["P{}".format(linea_archivo)].value = var3
#                                 worksheet["Q{}".format(linea_archivo)].value = 'SI' if var4 else ''
#
#                             if nummatriculas >= 3:
#                                 var5 = 'SI' if nummatriculas >= 3 else ''
#                                 var6 = qsmatriculas[2].materias_aprobadas_todas()
#
#                                 worksheet["R{}".format(linea_archivo)].value = var5
#                                 worksheet["S{}".format(linea_archivo)].value = 'SI' if var6 else ''
#
#                             qsprimernivel = Matricula.objects.filter(status=True, inscripcion__persona__cedula__icontains=cedula_, nivelmalla_id=1, inscripcion__carrera__nombre__unaccent__icontains=carrera_).exclude(inscripcion__coordinacion_id=9)
#                             if qsprimernivel.exists():
#                                 var7 = 'SI'
#                                 var8 = qsprimernivel.first().nivel.periodo.__str__()
#
#                                 worksheet["T{}".format(linea_archivo)].value = 'SI'
#                                 worksheet["U{}".format(linea_archivo)].value = var8
#
#                             worksheet["V{}".format(linea_archivo)].value = cedula_
#                         conexito += 1
#                     else:
#                         lis_excluidos.append({'cedula': cedula_, 'obs': 'No existe inscripción'})
#                         excluidos += 1
#                 else:
#                     lis_excluidos.append({'cedula': cedula_, 'obs': 'No existe persona'})
#                     excluidos += 1
#                 print("Linea {}/? - Persona: {}, Inscripciones: {}, Matricula: {} | {} - {} - {} - {} - {} - {} - {} - {}".format(linea, cedula_, numinscripciones, nummatriculas, var1, var2, var3, var4, var5, var6, var7, var8))
#             linea += 1
#             linea_archivo += 1
#             if linea in (25, 50, 400, 800, 1200, 3000, 5000, 6000, 8000, 10000, 15000, 18000, 20000):
#                 wb.save(url_archivo)
#                 print("Guardado Rapido Linea {} . .".format(linea))
#         print('Total Leidos con exito: {}'.format(conexito))
#         print('Total Excluidos: {}'.format(excluidos))
#         # print(lis_excluidos)
#         wb.save(url_archivo)
#         enviar_mensaje_bot_telegram('FINALIZO CON EXITO: {}'.format(url_archivo))
#     except Exception as ex:
#         textoerror = '{} Linea:{} - Info: {} / Lectura: {}'.format(str(ex), sys.exc_info()[-1].tb_lineno, cadena, linea)
#         print(textoerror)
#         enviar_mensaje_bot_telegram(textoerror)
#
#
# def matriz3():
#     try:
#         libre_origen = '/matriz_tabla_3_1.xls'
#         fuentecabecera = easyxf('font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
#         fuentenormal = easyxf('font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
#         output_folder = MEDIA_ROOT
#         output_folder = f'{SITE_STORAGE}/media/'
#         liborigen = xlrd.open_workbook(output_folder + libre_origen)
#         print(output_folder)
#         libdestino = xlwt.Workbook()
#         periodo_ = Periodo.objects.get(pk=126)
#         hojadestino = libdestino.add_sheet(f'{periodo_.nombre}')
#         fil = 0
#         columnas = [
#             (u"PERIODO", 7000),
#             (u"CARRERA", 2500),
#             (u"CEDULA", 2000),
#             (u"APELLIDO", 4000),
#             (u"NOMBRE", 4500),
#             (u"NIVEL", 9000),
#             (u"NIVEL SOCIO ECONOMICO", 4000),
#             (u"VALOR MATRICULA", 4000),
#             (u"SALDO", 4000),
#             (u"CREDITOS MALLA", 4000),
#             (u"TIPO MATRICULA", 4000),
#             (u"CODIGO MATRICULA", 4000),
#         ]
#         for col_num in range(len(columnas)):
#             hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
#             hojadestino.col(col_num).width = columnas[col_num][1]
#         fila = 1
#         qsbase = Matricula.objects.filter(nivel__periodo_id=periodo_.id, status=True,
#                                           cerrada=False, rubro__isnull=False, rubro__status=True)
#         segundamatriz = Matricula.objects.filter(id__in=qsbase.values_list('id', flat=True)).distinct().order_by('inscripcion__persona__apellido1', 'inscripcion__persona_apellido2')
#         print('Periodo: {} - Total: {}'.format(periodo_.__str__(), len(segundamatriz)))
#         for sm in segundamatriz:
#             nombre_ = "{} {} {}".format(sm.inscripcion.persona.apellido1, sm.inscripcion.persona.apellido2, sm.inscripcion.persona.nombres)
#             print('{}/{} - {}'.format(len(segundamatriz), fila, nombre_))
#             hojadestino.write(fila, 0, sm.nivel.periodo.nombre, fuentenormal)
#             hojadestino.write(fila, 1, sm.inscripcion.carrera.nombre, fuentenormal)
#             hojadestino.write(fila, 2, sm.inscripcion.persona.cedula, fuentenormal)
#             hojadestino.write(fila, 3, "{} {}".format(sm.inscripcion.persona.apellido1, sm.inscripcion.persona.apellido2), fuentenormal)
#             hojadestino.write(fila, 4, sm.inscripcion.persona.nombres, fuentenormal)
#             hojadestino.write(fila, 5, sm.nivelmalla.nombre, fuentenormal)
#             nombre_socio = ''
#             nivelsocioeco = ''
#             if sm.inscripcion.persona.fichasocioeconomicainec_set.exists():
#                 nombre_socio = '%s' % sm.inscripcion.persona.fichasocioeconomicainec_set.all()[0].grupoeconomico.nombre_corto()
#                 nivelsocioeco = '%s' % sm.inscripcion.persona.fichasocioeconomicainec_set.all()[0].grupoeconomico.id
#             hojadestino.write(fila, 6, nombre_socio, fuentenormal)
#             valor_matricula = Rubro.objects.filter(status=True, matricula=sm).aggregate(totalsum=Coalesce(Sum(('valortotal')), 0, output_field=FloatField())).get('totalsum')
#             saldo_matricula = Rubro.objects.filter(status=True, matricula=sm).aggregate(totalsum=Coalesce(Sum(('saldo')), 0, output_field=FloatField())).get('totalsum')
#             hojadestino.write(fila, 7, valor_matricula, fuentenormal)
#             hojadestino.write(fila, 8, saldo_matricula, fuentenormal)
#             hojadestino.write(fila, 9, sm.totalcreditos, fuentenormal)
#             hojadestino.write(fila, 10, sm.tipomatricula.nombre if sm.tipomatricula else '', fuentenormal)
#             hojadestino.write(fila, 11, sm.id, fuentenormal)
#             fila += 1
#
#         hojadestino = libdestino.add_sheet(f'{periodo_.nombre} RUBROS')
#         fil = 0
#         columnas = [
#             (u"CODIGO MATRICULA", 4000),
#             (u"NOMBRES", 2500),
#             (u"CUOTA", 2500),
#             (u"FECHA", 2500),
#             (u"NOMBRE", 4000),
#             (u"VALOR", 2000),
#             (u"SALDO", 2000),
#         ]
#         for col_num in range(len(columnas)):
#             hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
#             hojadestino.col(col_num).width = columnas[col_num][1]
#         fila = 1
#         for matricula in segundamatriz:
#             nombre_ = "{} {} {}".format(matricula.inscripcion.persona.apellido1, matricula.inscripcion.persona.apellido2, matricula.inscripcion.persona.nombres)
#             print('{}/{} - {}'.format(len(segundamatriz), fila, nombre_))
#             for sm in Rubro.objects.filter(status=True, matricula=matricula):
#                 hojadestino.write(fila, 0, sm.matricula.id, fuentenormal)
#                 hojadestino.write(fila, 1, nombre_, fuentenormal)
#                 hojadestino.write(fila, 2, sm.cuota, fuentenormal)
#                 hojadestino.write(fila, 3, str(sm.fecha), fuentenormal)
#                 hojadestino.write(fila, 4, sm.nombre, fuentenormal)
#                 hojadestino.write(fila, 5, sm.valortotal, fuentenormal)
#                 hojadestino.write(fila, 6, sm.saldo, fuentenormal)
#                 fila += 1
#
#         libdestino.save(output_folder + libre_origen)
#         print(output_folder + libre_origen)
#         print("Proceso finalizado. . .")
#     except Exception as ex:
#         transaction.set_rollback(True)
#         msg = ex.__str__()
#         print(msg)
#
#
# def matriz1():
#     try:
#         libre_origen = '/matriz_tabla_1.xls'
#         fuentecabecera = easyxf('font: name Verdana, color-index black, bold on, height 150; pattern: pattern solid, fore_colour gray25; alignment: vert distributed, horiz centre; borders: left thin, right thin, top thin, bottom thin')
#         fuentenormal = easyxf('font: name Verdana, color-index black, height 150; borders: left thin, right thin, top thin, bottom thin')
#         output_folder = MEDIA_ROOT
#         output_folder = os.path.join(os.path.join(BASE_DIR))
#         # liborigen = xlrd.open_workbook(output_folder + libre_origen)
#         libdestino = xlwt.Workbook()
#         periodo_ = Periodo.objects.get(pk=126)
#         hojadestino = libdestino.add_sheet(periodo_.nombre)
#         fil = 0
#         columnas = [
#             (u"CARRERA", 7000),
#             (u"NIVEL/SEMESTRE", 2500),
#             (u"NUMERO CREDITOS", 2000),
#         ]
#         for col_num in range(len(columnas)):
#             hojadestino.write(fil, col_num, columnas[col_num][0], fuentecabecera)
#             hojadestino.col(col_num).width = columnas[col_num][1]
#         fila = 1
#         segundamatriz = Clase.objects.filter(materia__nivel__periodo=periodo_, activo=True).distinct('materia')
#         print('Periodo: {} - Total: {}'.format(periodo_.__str__(), len(segundamatriz)))
#         for sm in segundamatriz:
#             print('{}/{} - {}'.format(len(segundamatriz), fila, sm))
#             malla = sm.materia.asignaturamalla.malla
#             hojadestino.write(fila, 0, sm.materia.asignaturamalla.malla.carrera.nombre, fuentenormal)
#             hojadestino.write(fila, 1, sm.materia.asignaturamalla.nivelmalla.nombre, fuentenormal)
#             hojadestino.write(fila, 2, sm.materia.asignaturamalla.nivelmalla.total_creditos(malla), fuentenormal)
#             fila += 1
#
#         libdestino.save(output_folder + libre_origen)
#         print(output_folder + libre_origen)
#         print("Proceso finalizado. . .")
#     except Exception as ex:
#         transaction.set_rollback(True)
#         msg = ex.__str__()
#         print(msg)
#
#
# def arreglo_frevision():
#     solicitudes = DatosEmpresaPreInscripcionPracticasPP.objects.filter(status=True).exclude(est_empresas=1)
#     for l in solicitudes:
#         print(l)
#         l.fecha_revision = l.fecha_notificacion
#         l.persona_revision = l.persona_notificacion
#         l.save()
#
#
# def subir_bitacora_personal_2():
#     workbook = xlrd.open_workbook("actividades_llerena.xlsx")
#     sheet = workbook.sheet_by_index(0)
#     linea = 1
#     col_fecha = 0
#     col_hora = 1
#     col_titulo = 2
#     col_actividad = 3
#     col_py_afectados = 4
#     departamento = Departamento.objects.get(pk=93)
#     persona = Persona.objects.get(pk=28164)
#     print(f"Se generaran {sheet.nrows - 1} registro de bitacora diairia a {persona.__str__()}")
#     for rowx in range(sheet.nrows):
#         if linea > 1:
#             print(f"Fila {linea} / {sheet.nrows}")
#             cols = sheet.row_values(rowx)
#             fecha = xlrd.xldate.xldate_as_datetime(cols[col_fecha], workbook.datemode)
#             hora = xlrd.xldate.xldate_as_datetime(cols[col_hora], workbook.datemode)
#             f = datetime(fecha.year, fecha.month, fecha.day, hora.hour, hora.minute, hora.second)
#             with transaction.atomic():
#                 try:
#                     if not BitacoraActividadDiaria.objects.filter(fecha=f).exists():
#                         eBitacoraActividadDiaria = BitacoraActividadDiaria(titulo=cols[col_titulo],
#                                                                            departamento=departamento,
#                                                                            fecha=f,
#                                                                            persona=persona,
#                                                                            descripcion=f"{cols[col_actividad]} en los archivos: {cols[col_py_afectados]}",
#                                                                            tiposistema=2)
#                         eBitacoraActividadDiaria.save(usuario_id=persona.usuario.id)
#                         print(f"Se guardo registro de fecha {eBitacoraActividadDiaria.fecha.__str__()}")
#                 except Exception as ex:
#                     transaction.set_rollback(True)
#                     print(f"No se guardo registro de fecha {f.__str__()}")
#         linea += 1


# def migrar_bitacora():
#     personas = [33490, 28164]
#     for p in personas:
#         persona = Persona.objects.get(id=p)
#         qsbase = IncidenciaSCRUM.objects.filter(asignadoa=persona)
#         for actividad in qsbase:
#             if not BitacoraActividadDiaria.objects.filter(status=True, fecha=actividad.finicioactividad,
#                                                           descripcion=actividad.descripcion).exists():
#                 bitacora = BitacoraActividadDiaria(persona=persona,
#                                                    fecha=actividad.finicioactividad,
#                                                    fechafin=actividad.ffinactividad,
#                                                    departamento=persona.mi_departamento(),
#                                                    descripcion=actividad.descripcion,
#                                                    link='',
#                                                    tiposistema=2,
#                                                    departamento_requiriente=actividad.categoria.direccion)
#                 bitacora.save()
#
# migrar_bitacora()


# def migrar_email_inst():
#     qsper = Persona.objects.filter(status=True, usuario__isnull=False, emailinst__isnull=False)
#     tot = 1
#     total_ = qsper.count()
#     for per in qsper:
#         emailinst = per.emailinst
#         print(f"{total_}/{tot}) {per} - {emailinst}")
#         user_ = User.objects.get(id=per.usuario_id)
#         user_.email = emailinst
#         user_.save()
#         tot += 1
#
# migrar_email_inst()

#
# def arreglo_solicitud_balcon():
#     try:
#         from balcon.models import Solicitud
#         qs = Solicitud.objects.filter(status=True)
#         total_qs, count = qs.count(), 1
#         for ins in qs:
#             print(f"{count}/{total_qs} {ins}")
#             ins.agenteactual = ins.ver_servicio().asignadorecibe
#             ins.save()
#             count += 1
#     except Exception as ex:
#         print(ex)
#
# arreglo_solicitud_balcon()

# from voto.models import *
#
# def sedeelectoral():
#     try:
#         # Obtener el registro más reciente por persona_id
#         registros_mas_recientes = (
#             PersonasSede.objects.filter(
#                 sede_id__in=[12, 13],
#                 persona_id__in=[
#                     49148, 96845, 54569, 97391, 89452, 20972, 95796, 90516, 59280, 118291, 72445,
#                     52492, 96136, 57472, 53519, 93110, 46488, 98116, 128787, 109187, 103272, 106606,
#                     39425, 67390, 43509, 52309, 80055, 82402, 101360, 116154, 75624, 100964, 70837,
#                     58473, 102570, 49183, 68433, 69852, 88962, 10680, 66539, 72795, 72783, 53041,
#                     108116, 97642, 107740, 52146, 84518, 119568, 85809, 43746, 45625, 74100, 62136,
#                     90605, 87732, 89027, 128698, 135681, 96894, 99041, 119268, 29500, 45199, 79110,
#                     76666, 26030, 111586, 49925, 109499, 95686, 15362, 78131, 86861, 51968, 108460,
#                     116394, 106586, 91506, 113436, 72750
#                 ],
#             )
#                 .values('persona_id')
#                 .annotate(ultima_fecha=Max('fecha_creacion'))
#         )
#
#         # Actualizar registros para desactivar aquellos que no son los más recientes
#         for registro in registros_mas_recientes:
#             persona_id = registro['persona_id']
#             ultima_fecha = registro['ultima_fecha']
#
#             # Desactivar registros anteriores
#             PersonasSede.objects.filter(persona_id=persona_id, sede_id__in=[12, 13], fecha_creacion__lt=ultima_fecha).update(status=False)
#
#             # Activar el registro más reciente
#             PersonasSede.objects.filter(persona_id=persona_id, sede_id__in=[12, 13], fecha_creacion=ultima_fecha).update(status=True)
#
#
#
#     except Exception as ex:
#         print(ex)
#

#
# def consumo_ws_sen():
#     try:
#         import requests
#         import xml.etree.ElementTree as ET
#
#         xml_data = '''<soap:Envelope xmlns:soap="http://www.w3.org/2003/05/soap-envelope" xmlns:swr="http://sw.registrounicoedusup.gob.ec:8080/SW-SNNA/SWRegistroNacional">
#            <soap:Header/>
#                <soap:Body>
#                   <swr:consultarRegistroNacional>
#                      <codigoIes></codigoIes>
#                      <identificacion>0928541556</identificacion>
#                   </swr:consultarRegistroNacional>
#                </soap:Body>
#             </soap:Envelope>'''
#
#         # URL del servicio web
#         url = 'http://sw.registrounicoedusup.gob.ec:8080/SW-SNNA/SWRegistroNacional?wsdl'
#
#         # Encabezados para la solicitud
#         headers = {
#             'Content-Type': 'application/xml'
#         }
#
#         # Realizar la solicitud POST al servicio web
#         response = requests.post(url, data=xml_data, headers=headers)
#
#         # Verificar si la solicitud fue exitosa (código de estado 200)
#         if response.status_code == 200:
#             # Analizar la respuesta XML
#             print(response.text)
#             # Realizar operaciones con los datos XML recibidos
#         else:
#             print(f"Error en la solicitud. Código de estado: {response.status_code}")
#
#
#     except Exception as ex:
#         print(ex)

# consumo_ws_sen()



# from voto.models import *
#
# def sedeelectoral():
#     try:
#         qspersonas = DetPersonaPadronElectoral.objects.filter(status=True, tipo=1, cab__id=3).order_by('id')
#         total_, count = len(qspersonas), 1
#         for qp in qspersonas:
#             print(f"{count}/{total_} - {qp}")
#             seleccionsede_ = PersonasSede.objects.filter(status=True, persona=qp.persona, sede__periodo__id=3).order_by('-id')
#             if seleccionsede_:
#                 sede_ = seleccionsede_.first().sede
#                 qp.lugarsede = sede_
#             else:
#                 qp.lugarsede_id = 12
#             qp.save()
#             count += 1
#     except Exception as ex:
#         print(ex)
#
# sedeelectoral()
#
#
# asunto = u"INSIGNIA - "
# send_html_mail(asunto, "emails/notificar_tituloinsignia_posgrado.html",
#                {'sistema': 'sga'},
#                'hllerenaa@unemi.edu.ec',
#                [], None,
#                cuenta=CUENTAS_CORREOS[0][1])



# def empadronarpersonas2023sd():
#     cadena = ''
#     linea, excluidos, conexito = 0, 0, 0
#     try:
#         archivo_ = 'SD_MILAGRO'
#         # url_archivo = "{}/{}.xlsx".format(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), archivo_)
#         url_archivo = "{}/media/{}.xlsx".format(SITE_STORAGE, archivo_)
#         print('Leyendo....')
#         # wb = openpyxl.load_workbook(filename="{}/media/{}.xlsx".format(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), archivo_))
#         wb = openpyxl.load_workbook(filename=url_archivo)
#         ws = wb.get_sheet_by_name("Hoja1")
#         # worksheet = wb["Listado"]
#         worksheet = ws
#         lis_excluidos = []
#         print('Iniciando....')
#         linea_archivo = 1
#         ids_excluir_a_rechazar = []
#         periodo = CabPadronElectoral.objects.get(pk=3, status=True)
#         for row in worksheet.iter_rows(min_row=0):
#             if linea >= 1:
#                 currentValues, cadena = [], ''
#                 for cell in row:
#                     cadena += str(cell.value) + ' '
#                     currentValues.append(str(cell.value))
#                 cedula_ = currentValues[1]
#                 coordinacion = currentValues[3]
#                 mesa = currentValues[4]
#                 nombre_mesa = f"{mesa} {coordinacion} - SANTO DOMINGO"
#                 qsmesa = MesasPadronElectoral.objects.filter(status=True, periodo=periodo, orden=1, nombre=nombre_mesa)
#                 mesa_ = None
#                 if qsmesa.exists():
#                     mesa_ = qsmesa.first()
#                 else:
#                     mesa_ = MesasPadronElectoral(periodo=periodo, orden=1, nombre=nombre_mesa)
#                     mesa_.save()
#
#                 if mesa_:
#                     qsempadronada = DetPersonaPadronElectoral.objects.filter(status=True, cab=periodo, persona__cedula__icontains=cedula_, tipo=1)
#                     empadronado = None
#                     if qsempadronada.exists():
#                         empadronado = qsempadronada.first()
#                         empadronado.lugar = nombre_mesa
#                         empadronado.mesa = mesa_
#                         empadronado.save()
#                     else:
#                         lis_excluidos.append({'cedula': cedula_, 'obs': f'No existe empadronado'})
#                         excluidos += 1
#                 print("Linea {}/? - Persona: {}".format(linea, cedula_))
#             linea += 1
#             linea_archivo += 1
#         print('Total Leidos con exito: {}'.format(conexito))
#         print('Total Excluidos: {}'.format(excluidos))
#         print(lis_excluidos)
#     except Exception as ex:
#         textoerror = '{} Linea:{} - Info: {} / Lectura: {}'.format(str(ex), sys.exc_info()[-1].tb_lineno, cadena, linea)
#         print(textoerror)


# def empadronarpersonas2023admd0c():
#     cadena = ''
#     linea, excluidos, conexito = 0, 0, 0
#     try:
#         archivo_ = 'TRABAJADORESDOCENTES'
#         # url_archivo = "{}/{}.xlsx".format(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), archivo_)
#         url_archivo = "{}/media/{}.xlsx".format(SITE_STORAGE, archivo_)
#         print('Leyendo....')
#         # wb = openpyxl.load_workbook(filename="{}/media/{}.xlsx".format(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), archivo_))
#         wb = openpyxl.load_workbook(filename=url_archivo)
#         ws = wb.get_sheet_by_name("Hoja2")
#         # worksheet = wb["Listado"]
#         worksheet = ws
#         lis_excluidos = []
#         print('Iniciando....')
#         linea_archivo = 1
#         ids_excluir_a_rechazar = []
#         periodo = CabPadronElectoral.objects.get(pk=3, status=True)
#         for row in worksheet.iter_rows(min_row=0):
#             if linea >= 1:
#                 currentValues, cadena = [], ''
#                 for cell in row:
#                     cadena += str(cell.value) + ' '
#                     currentValues.append(str(cell.value))
#                 cedula_ = currentValues[1]
#                 coordinacion = currentValues[3]
#                 mesa = currentValues[4]
#                 nombre_mesa = f"{mesa} {coordinacion} - MILAGRO"
#                 qsmesa = MesasPadronElectoral.objects.filter(status=True, periodo=periodo, orden=1, nombre=nombre_mesa)
#                 mesa_ = None
#                 if qsmesa.exists():
#                     mesa_ = qsmesa.first()
#                 else:
#                     mesa_ = MesasPadronElectoral(periodo=periodo, orden=1, nombre=nombre_mesa)
#                     mesa_.save()
#
#                 if mesa_:
#                     qsempadronada = DetPersonaPadronElectoral.objects.filter(status=True, cab=periodo, persona__cedula__icontains=cedula_, tipo=2)
#                     empadronado = None
#                     if qsempadronada.exists():
#                         empadronado = qsempadronada.first()
#                         empadronado.lugar = nombre_mesa
#                         empadronado.mesa = mesa_
#                         empadronado.save()
#                     else:
#                         lis_excluidos.append({'cedula': cedula_, 'obs': f'No existe empadronado'})
#                         excluidos += 1
#                 print("Linea {}/? - Persona: {}".format(linea, cedula_))
#             linea += 1
#             linea_archivo += 1
#         print('Total Leidos con exito: {}'.format(conexito))
#         print('Total Excluidos: {}'.format(excluidos))
#         print(lis_excluidos)
#     except Exception as ex:
#         textoerror = '{} Linea:{} - Info: {} / Lectura: {}'.format(str(ex), sys.exc_info()[-1].tb_lineno, cadena, linea)
#         print(textoerror)
#
#
# empadronarpersonas2023admd0c()


def procesartotalmesas():
    from django.db.models import Max, F
    from django.db.models import CharField, Value
    from django.db.models.functions import Concat
    from sga.models import Persona, FotoPersona

    # Obtener todas las personas con su última foto
    personas_con_ultima_foto = Persona.objects.annotate(
        ultima_foto=Max('fotopersona__id')
    ).values(
        'nombres',
        'apellido1',
        'apellido2',
        'cedula',
        'pasaporte',
        'ruc',
        'tipopersona',
        'contribuyenteespecial',
        'nacimiento',
        'sexo__nombre',  # Asumiendo que Sexo tiene un campo "nombre"
        'anioresidencia',
        'paisnacimiento__nombre',  # Asumiendo que Pais tiene un campo "nombre"
        'provincianacimiento__nombre',  # Asumiendo que Provincia tiene un campo "nombre"
        'cantonnacimiento__nombre',  # Asumiendo que Canton tiene un campo "nombre"
        'parroquianacimiento__nombre',  # Asumiendo que Parroquia tiene un campo "nombre"
        'nacionalidad',
        'paisnacionalidad__nombre',  # Asumiendo que Pais tiene un campo "nombre"
        'pais__nombre',  # Asumiendo que Pais tiene un campo "nombre"
        'provincia__nombre',  # Asumiendo que Provincia tiene un campo "nombre"
        'canton__nombre',  # Asumiendo que Canton tiene un campo "nombre"
        'parroquia__nombre',  # Asumiendo que Parroquia tiene un campo "nombre"
        'ciudadela',
        'sector',
        'ciudad',
        'direccion',
        'direccion2',
        'num_direccion',
        'referencia',
        'telefono',
        'telefono_conv',
        'email',
        'emailinst',
        ultima_foto=F('fotopersona__foto')
    )

    # Ahora tienes un queryset con todas las personas y su última foto como "ultima_foto".
    # Puedes iterar sobre este queryset para acceder a los valores de cada persona y su última foto.
    for persona in personas_con_ultima_foto:
        print(persona)

    # for confmesa in ConfiguracionMesaResponsable.objects.filter(status=True, periodo__id=3):
    #     print(f"{confmesa}")
    #     confmesa.totalempadronados = confmesa.mesa.detpersonas()
    #     confmesa.save()
    #     for detmesa in confmesa.detallemesa_set.filter(status=True):
    #         print(f"----{detmesa}")
    #         detmesa.empadronado = confmesa.mesa.detpersonas()
    #         detmesa.save()

procesartotalmesas()